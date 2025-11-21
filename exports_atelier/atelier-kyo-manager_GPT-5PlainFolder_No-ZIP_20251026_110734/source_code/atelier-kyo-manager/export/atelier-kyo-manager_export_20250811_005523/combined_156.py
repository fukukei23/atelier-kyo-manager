
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\commands\lock.py ===
import sys
from optparse import Values
from pathlib import Path
from typing import List

from pip._internal.cache import WheelCache
from pip._internal.cli import cmdoptions
from pip._internal.cli.req_command import (
    RequirementCommand,
    with_cleanup,
)
from pip._internal.cli.status_codes import SUCCESS
from pip._internal.models.pylock import Pylock, is_valid_pylock_file_name
from pip._internal.operations.build.build_tracker import get_build_tracker
from pip._internal.req.req_install import (
    check_legacy_setup_py_options,
)
from pip._internal.utils.logging import getLogger
from pip._internal.utils.misc import (
    get_pip_version,
)
from pip._internal.utils.temp_dir import TempDirectory

logger = getLogger(__name__)


class LockCommand(RequirementCommand):
    """
    EXPERIMENTAL - Lock packages and their dependencies from:

    - PyPI (and other indexes) using requirement specifiers.
    - VCS project urls.
    - Local project directories.
    - Local or remote source archives.

    pip also supports locking from "requirements files", which provide an easy
    way to specify a whole environment to be installed.

    The generated lock file is only guaranteed to be valid for the current
    python version and platform.
    """

    usage = """
      %prog [options] [-e] <local project path> ...
      %prog [options] <requirement specifier> [package-index-options] ...
      %prog [options] -r <requirements file> [package-index-options] ...
      %prog [options] <archive url/path> ..."""

    def add_options(self) -> None:
        self.cmd_opts.add_option(
            cmdoptions.PipOption(
                "--output",
                "-o",
                dest="output_file",
                metavar="path",
                type="path",
                default="pylock.toml",
                help="Lock file name (default=pylock.toml). Use - for stdout.",
            )
        )
        self.cmd_opts.add_option(cmdoptions.requirements())
        self.cmd_opts.add_option(cmdoptions.constraints())
        self.cmd_opts.add_option(cmdoptions.no_deps())
        self.cmd_opts.add_option(cmdoptions.pre())

        self.cmd_opts.add_option(cmdoptions.editable())

        self.cmd_opts.add_option(cmdoptions.src())

        self.cmd_opts.add_option(cmdoptions.ignore_requires_python())
        self.cmd_opts.add_option(cmdoptions.no_build_isolation())
        self.cmd_opts.add_option(cmdoptions.use_pep517())
        self.cmd_opts.add_option(cmdoptions.no_use_pep517())
        self.cmd_opts.add_option(cmdoptions.check_build_deps())

        self.cmd_opts.add_option(cmdoptions.config_settings())

        self.cmd_opts.add_option(cmdoptions.no_binary())
        self.cmd_opts.add_option(cmdoptions.only_binary())
        self.cmd_opts.add_option(cmdoptions.prefer_binary())
        self.cmd_opts.add_option(cmdoptions.require_hashes())
        self.cmd_opts.add_option(cmdoptions.progress_bar())

        index_opts = cmdoptions.make_option_group(
            cmdoptions.index_group,
            self.parser,
        )

        self.parser.insert_option_group(0, index_opts)
        self.parser.insert_option_group(0, self.cmd_opts)

    @with_cleanup
    def run(self, options: Values, args: List[str]) -> int:
        logger.verbose("Using %s", get_pip_version())

        logger.warning(
            "pip lock is currently an experimental command. "
            "It may be removed/changed in a future release "
            "without prior warning."
        )

        session = self.get_default_session(options)

        finder = self._build_package_finder(
            options=options,
            session=session,
            ignore_requires_python=options.ignore_requires_python,
        )
        build_tracker = self.enter_context(get_build_tracker())

        directory = TempDirectory(
            delete=not options.no_clean,
            kind="install",
            globally_managed=True,
        )

        reqs = self.get_requirements(args, options, finder, session)
        check_legacy_setup_py_options(options, reqs)

        wheel_cache = WheelCache(options.cache_dir)

        # Only when installing is it permitted to use PEP 660.
        # In other circumstances (pip wheel, pip download) we generate
        # regular (i.e. non editable) metadata and wheels.
        for req in reqs:
            req.permit_editable_wheels = True

        preparer = self.make_requirement_preparer(
            temp_build_dir=directory,
            options=options,
            build_tracker=build_tracker,
            session=session,
            finder=finder,
            use_user_site=False,
            verbosity=self.verbosity,
        )
        resolver = self.make_resolver(
            preparer=preparer,
            finder=finder,
            options=options,
            wheel_cache=wheel_cache,
            use_user_site=False,
            ignore_installed=True,
            ignore_requires_python=options.ignore_requires_python,
            upgrade_strategy="to-satisfy-only",
            use_pep517=options.use_pep517,
        )

        self.trace_basic_info(finder)

        requirement_set = resolver.resolve(reqs, check_supported_wheels=True)

        if options.output_file == "-":
            base_dir = Path.cwd()
        else:
            output_file_path = Path(options.output_file)
            if not is_valid_pylock_file_name(output_file_path):
                logger.warning(
                    "%s is not a valid lock file name.",
                    output_file_path,
                )
            base_dir = output_file_path.parent
        pylock_toml = Pylock.from_install_requirements(
            requirement_set.requirements.values(), base_dir=base_dir
        ).as_toml()
        if options.output_file == "-":
            sys.stdout.write(pylock_toml)
        else:
            output_file_path.write_text(pylock_toml, encoding="utf-8")

        return SUCCESS

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\flask\json\__init__.py ===
from __future__ import annotations

import json as _json
import typing as t

from ..globals import current_app
from .provider import _default

if t.TYPE_CHECKING:  # pragma: no cover
    from ..wrappers import Response


def dumps(obj: t.Any, **kwargs: t.Any) -> str:
    """Serialize data as JSON.

    If :data:`~flask.current_app` is available, it will use its
    :meth:`app.json.dumps() <flask.json.provider.JSONProvider.dumps>`
    method, otherwise it will use :func:`json.dumps`.

    :param obj: The data to serialize.
    :param kwargs: Arguments passed to the ``dumps`` implementation.

    .. versionchanged:: 2.3
        The ``app`` parameter was removed.

    .. versionchanged:: 2.2
        Calls ``current_app.json.dumps``, allowing an app to override
        the behavior.

    .. versionchanged:: 2.0.2
        :class:`decimal.Decimal` is supported by converting to a string.

    .. versionchanged:: 2.0
        ``encoding`` will be removed in Flask 2.1.

    .. versionchanged:: 1.0.3
        ``app`` can be passed directly, rather than requiring an app
        context for configuration.
    """
    if current_app:
        return current_app.json.dumps(obj, **kwargs)

    kwargs.setdefault("default", _default)
    return _json.dumps(obj, **kwargs)


def dump(obj: t.Any, fp: t.IO[str], **kwargs: t.Any) -> None:
    """Serialize data as JSON and write to a file.

    If :data:`~flask.current_app` is available, it will use its
    :meth:`app.json.dump() <flask.json.provider.JSONProvider.dump>`
    method, otherwise it will use :func:`json.dump`.

    :param obj: The data to serialize.
    :param fp: A file opened for writing text. Should use the UTF-8
        encoding to be valid JSON.
    :param kwargs: Arguments passed to the ``dump`` implementation.

    .. versionchanged:: 2.3
        The ``app`` parameter was removed.

    .. versionchanged:: 2.2
        Calls ``current_app.json.dump``, allowing an app to override
        the behavior.

    .. versionchanged:: 2.0
        Writing to a binary file, and the ``encoding`` argument, will be
        removed in Flask 2.1.
    """
    if current_app:
        current_app.json.dump(obj, fp, **kwargs)
    else:
        kwargs.setdefault("default", _default)
        _json.dump(obj, fp, **kwargs)


def loads(s: str | bytes, **kwargs: t.Any) -> t.Any:
    """Deserialize data as JSON.

    If :data:`~flask.current_app` is available, it will use its
    :meth:`app.json.loads() <flask.json.provider.JSONProvider.loads>`
    method, otherwise it will use :func:`json.loads`.

    :param s: Text or UTF-8 bytes.
    :param kwargs: Arguments passed to the ``loads`` implementation.

    .. versionchanged:: 2.3
        The ``app`` parameter was removed.

    .. versionchanged:: 2.2
        Calls ``current_app.json.loads``, allowing an app to override
        the behavior.

    .. versionchanged:: 2.0
        ``encoding`` will be removed in Flask 2.1. The data must be a
        string or UTF-8 bytes.

    .. versionchanged:: 1.0.3
        ``app`` can be passed directly, rather than requiring an app
        context for configuration.
    """
    if current_app:
        return current_app.json.loads(s, **kwargs)

    return _json.loads(s, **kwargs)


def load(fp: t.IO[t.AnyStr], **kwargs: t.Any) -> t.Any:
    """Deserialize data as JSON read from a file.

    If :data:`~flask.current_app` is available, it will use its
    :meth:`app.json.load() <flask.json.provider.JSONProvider.load>`
    method, otherwise it will use :func:`json.load`.

    :param fp: A file opened for reading text or UTF-8 bytes.
    :param kwargs: Arguments passed to the ``load`` implementation.

    .. versionchanged:: 2.3
        The ``app`` parameter was removed.

    .. versionchanged:: 2.2
        Calls ``current_app.json.load``, allowing an app to override
        the behavior.

    .. versionchanged:: 2.2
        The ``app`` parameter will be removed in Flask 2.3.

    .. versionchanged:: 2.0
        ``encoding`` will be removed in Flask 2.1. The file must be text
        mode, or binary mode with UTF-8 bytes.
    """
    if current_app:
        return current_app.json.load(fp, **kwargs)

    return _json.load(fp, **kwargs)


def jsonify(*args: t.Any, **kwargs: t.Any) -> Response:
    """Serialize the given arguments as JSON, and return a
    :class:`~flask.Response` object with the ``application/json``
    mimetype. A dict or list returned from a view will be converted to a
    JSON response automatically without needing to call this.

    This requires an active request or application context, and calls
    :meth:`app.json.response() <flask.json.provider.JSONProvider.response>`.

    In debug mode, the output is formatted with indentation to make it
    easier to read. This may also be controlled by the provider.

    Either positional or keyword arguments can be given, not both.
    If no arguments are given, ``None`` is serialized.

    :param args: A single value to serialize, or multiple values to
        treat as a list to serialize.
    :param kwargs: Treat as a dict to serialize.

    .. versionchanged:: 2.2
        Calls ``current_app.json.response``, allowing an app to override
        the behavior.

    .. versionchanged:: 2.0.2
        :class:`decimal.Decimal` is supported by converting to a string.

    .. versionchanged:: 0.11
        Added support for serializing top-level arrays. This was a
        security risk in ancient browsers. See :ref:`security-json`.

    .. versionadded:: 0.2
    """
    return current_app.json.response(*args, **kwargs)  # type: ignore[return-value]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\__config__.py ===
# This file is generated by numpy's build process
# It contains system_info results at the time of building this package.
from enum import Enum
from numpy._core._multiarray_umath import (
    __cpu_features__,
    __cpu_baseline__,
    __cpu_dispatch__,
)

__all__ = ["show_config"]
_built_with_meson = True


class DisplayModes(Enum):
    stdout = "stdout"
    dicts = "dicts"


def _cleanup(d):
    """
    Removes empty values in a `dict` recursively
    This ensures we remove values that Meson could not provide to CONFIG
    """
    if isinstance(d, dict):
        return {k: _cleanup(v) for k, v in d.items() if v and _cleanup(v)}
    else:
        return d


CONFIG = _cleanup(
    {
        "Compilers": {
            "c": {
                "name": "msvc",
                "linker": r"link",
                "version": "19.43.34808",
                "commands": r"cl",
                "args": r"",
                "linker args": r"",
            },
            "cython": {
                "name": "cython",
                "linker": r"cython",
                "version": "3.1.1",
                "commands": r"cython",
                "args": r"",
                "linker args": r"",
            },
            "c++": {
                "name": "msvc",
                "linker": r"link",
                "version": "19.43.34808",
                "commands": r"cl",
                "args": r"",
                "linker args": r"",
            },
        },
        "Machine Information": {
            "host": {
                "cpu": "x86_64",
                "family": "x86_64",
                "endian": "little",
                "system": "windows",
            },
            "build": {
                "cpu": "x86_64",
                "family": "x86_64",
                "endian": "little",
                "system": "windows",
            },
            "cross-compiled": bool("False".lower().replace("false", "")),
        },
        "Build Dependencies": {
            "blas": {
                "name": "scipy-openblas",
                "found": bool("True".lower().replace("false", "")),
                "version": "0.3.29",
                "detection method": "pkgconfig",
                "include directory": r"C:/Users/runneradmin/AppData/Local/Temp/cibw-run-ly2cw5el/cp312-win_amd64/build/venv/Lib/site-packages/scipy_openblas64/include",
                "lib directory": r"C:/Users/runneradmin/AppData/Local/Temp/cibw-run-ly2cw5el/cp312-win_amd64/build/venv/Lib/site-packages/scipy_openblas64/lib",
                "openblas configuration": r"OpenBLAS 0.3.29  USE64BITINT DYNAMIC_ARCH NO_AFFINITY Haswell MAX_THREADS=24",
                "pc file directory": r"C:/a/numpy/numpy/.openblas",
            },
            "lapack": {
                "name": "scipy-openblas",
                "found": bool("True".lower().replace("false", "")),
                "version": "0.3.29",
                "detection method": "pkgconfig",
                "include directory": r"C:/Users/runneradmin/AppData/Local/Temp/cibw-run-ly2cw5el/cp312-win_amd64/build/venv/Lib/site-packages/scipy_openblas64/include",
                "lib directory": r"C:/Users/runneradmin/AppData/Local/Temp/cibw-run-ly2cw5el/cp312-win_amd64/build/venv/Lib/site-packages/scipy_openblas64/lib",
                "openblas configuration": r"OpenBLAS 0.3.29  USE64BITINT DYNAMIC_ARCH NO_AFFINITY Haswell MAX_THREADS=24",
                "pc file directory": r"C:/a/numpy/numpy/.openblas",
            },
        },
        "Python Information": {
            "path": r"C:\Users\runneradmin\AppData\Local\Temp\build-env-5mlc73_a\Scripts\python.exe",
            "version": "3.12",
        },
        "SIMD Extensions": {
            "baseline": __cpu_baseline__,
            "found": [
                feature for feature in __cpu_dispatch__ if __cpu_features__[feature]
            ],
            "not found": [
                feature for feature in __cpu_dispatch__ if not __cpu_features__[feature]
            ],
        },
    }
)


def _check_pyyaml():
    import yaml

    return yaml


def show(mode=DisplayModes.stdout.value):
    """
    Show libraries and system information on which NumPy was built
    and is being used

    Parameters
    ----------
    mode : {`'stdout'`, `'dicts'`}, optional.
        Indicates how to display the config information.
        `'stdout'` prints to console, `'dicts'` returns a dictionary
        of the configuration.

    Returns
    -------
    out : {`dict`, `None`}
        If mode is `'dicts'`, a dict is returned, else None

    See Also
    --------
    get_include : Returns the directory containing NumPy C
                  header files.

    Notes
    -----
    1. The `'stdout'` mode will give more readable
       output if ``pyyaml`` is installed

    """
    if mode == DisplayModes.stdout.value:
        try:  # Non-standard library, check import
            yaml = _check_pyyaml()

            print(yaml.dump(CONFIG))
        except ModuleNotFoundError:
            import warnings
            import json

            warnings.warn("Install `pyyaml` for better output", stacklevel=1)
            print(json.dumps(CONFIG, indent=2))
    elif mode == DisplayModes.dicts.value:
        return CONFIG
    else:
        raise AttributeError(
            f"Invalid `mode`, use one of: {', '.join([e.value for e in DisplayModes])}"
        )


def show_config(mode=DisplayModes.stdout.value):
    return show(mode)


show_config.__doc__ = show.__doc__
show_config.__module__ = "numpy"

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\_testing\_io.py ===
from __future__ import annotations

import gzip
import io
import pathlib
import tarfile
from typing import (
    TYPE_CHECKING,
    Any,
    Callable,
)
import uuid
import zipfile

from pandas.compat import (
    get_bz2_file,
    get_lzma_file,
)
from pandas.compat._optional import import_optional_dependency

import pandas as pd
from pandas._testing.contexts import ensure_clean

if TYPE_CHECKING:
    from pandas._typing import (
        FilePath,
        ReadPickleBuffer,
    )

    from pandas import (
        DataFrame,
        Series,
    )

# ------------------------------------------------------------------
# File-IO


def round_trip_pickle(
    obj: Any, path: FilePath | ReadPickleBuffer | None = None
) -> DataFrame | Series:
    """
    Pickle an object and then read it again.

    Parameters
    ----------
    obj : any object
        The object to pickle and then re-read.
    path : str, path object or file-like object, default None
        The path where the pickled object is written and then read.

    Returns
    -------
    pandas object
        The original object that was pickled and then re-read.
    """
    _path = path
    if _path is None:
        _path = f"__{uuid.uuid4()}__.pickle"
    with ensure_clean(_path) as temp_path:
        pd.to_pickle(obj, temp_path)
        return pd.read_pickle(temp_path)


def round_trip_pathlib(writer, reader, path: str | None = None):
    """
    Write an object to file specified by a pathlib.Path and read it back

    Parameters
    ----------
    writer : callable bound to pandas object
        IO writing function (e.g. DataFrame.to_csv )
    reader : callable
        IO reading function (e.g. pd.read_csv )
    path : str, default None
        The path where the object is written and then read.

    Returns
    -------
    pandas object
        The original object that was serialized and then re-read.
    """
    Path = pathlib.Path
    if path is None:
        path = "___pathlib___"
    with ensure_clean(path) as path:
        writer(Path(path))  # type: ignore[arg-type]
        obj = reader(Path(path))  # type: ignore[arg-type]
    return obj


def round_trip_localpath(writer, reader, path: str | None = None):
    """
    Write an object to file specified by a py.path LocalPath and read it back.

    Parameters
    ----------
    writer : callable bound to pandas object
        IO writing function (e.g. DataFrame.to_csv )
    reader : callable
        IO reading function (e.g. pd.read_csv )
    path : str, default None
        The path where the object is written and then read.

    Returns
    -------
    pandas object
        The original object that was serialized and then re-read.
    """
    import pytest

    LocalPath = pytest.importorskip("py.path").local
    if path is None:
        path = "___localpath___"
    with ensure_clean(path) as path:
        writer(LocalPath(path))
        obj = reader(LocalPath(path))
    return obj


def write_to_compressed(compression, path, data, dest: str = "test") -> None:
    """
    Write data to a compressed file.

    Parameters
    ----------
    compression : {'gzip', 'bz2', 'zip', 'xz', 'zstd'}
        The compression type to use.
    path : str
        The file path to write the data.
    data : str
        The data to write.
    dest : str, default "test"
        The destination file (for ZIP only)

    Raises
    ------
    ValueError : An invalid compression value was passed in.
    """
    args: tuple[Any, ...] = (data,)
    mode = "wb"
    method = "write"
    compress_method: Callable

    if compression == "zip":
        compress_method = zipfile.ZipFile
        mode = "w"
        args = (dest, data)
        method = "writestr"
    elif compression == "tar":
        compress_method = tarfile.TarFile
        mode = "w"
        file = tarfile.TarInfo(name=dest)
        bytes = io.BytesIO(data)
        file.size = len(data)
        args = (file, bytes)
        method = "addfile"
    elif compression == "gzip":
        compress_method = gzip.GzipFile
    elif compression == "bz2":
        compress_method = get_bz2_file()
    elif compression == "zstd":
        compress_method = import_optional_dependency("zstandard").open
    elif compression == "xz":
        compress_method = get_lzma_file()
    else:
        raise ValueError(f"Unrecognized compression type: {compression}")

    with compress_method(path, mode=mode) as f:
        getattr(f, method)(*args)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\msgpack\ext.py ===
import datetime
import struct
from collections import namedtuple


class ExtType(namedtuple("ExtType", "code data")):
    """ExtType represents ext type in msgpack."""

    def __new__(cls, code, data):
        if not isinstance(code, int):
            raise TypeError("code must be int")
        if not isinstance(data, bytes):
            raise TypeError("data must be bytes")
        if not 0 <= code <= 127:
            raise ValueError("code must be 0~127")
        return super().__new__(cls, code, data)


class Timestamp:
    """Timestamp represents the Timestamp extension type in msgpack.

    When built with Cython, msgpack uses C methods to pack and unpack `Timestamp`.
    When using pure-Python msgpack, :func:`to_bytes` and :func:`from_bytes` are used to pack and
    unpack `Timestamp`.

    This class is immutable: Do not override seconds and nanoseconds.
    """

    __slots__ = ["seconds", "nanoseconds"]

    def __init__(self, seconds, nanoseconds=0):
        """Initialize a Timestamp object.

        :param int seconds:
            Number of seconds since the UNIX epoch (00:00:00 UTC Jan 1 1970, minus leap seconds).
            May be negative.

        :param int nanoseconds:
            Number of nanoseconds to add to `seconds` to get fractional time.
            Maximum is 999_999_999.  Default is 0.

        Note: Negative times (before the UNIX epoch) are represented as neg. seconds + pos. ns.
        """
        if not isinstance(seconds, int):
            raise TypeError("seconds must be an integer")
        if not isinstance(nanoseconds, int):
            raise TypeError("nanoseconds must be an integer")
        if not (0 <= nanoseconds < 10**9):
            raise ValueError("nanoseconds must be a non-negative integer less than 999999999.")
        self.seconds = seconds
        self.nanoseconds = nanoseconds

    def __repr__(self):
        """String representation of Timestamp."""
        return f"Timestamp(seconds={self.seconds}, nanoseconds={self.nanoseconds})"

    def __eq__(self, other):
        """Check for equality with another Timestamp object"""
        if type(other) is self.__class__:
            return self.seconds == other.seconds and self.nanoseconds == other.nanoseconds
        return False

    def __ne__(self, other):
        """not-equals method (see :func:`__eq__()`)"""
        return not self.__eq__(other)

    def __hash__(self):
        return hash((self.seconds, self.nanoseconds))

    @staticmethod
    def from_bytes(b):
        """Unpack bytes into a `Timestamp` object.

        Used for pure-Python msgpack unpacking.

        :param b: Payload from msgpack ext message with code -1
        :type b: bytes

        :returns: Timestamp object unpacked from msgpack ext payload
        :rtype: Timestamp
        """
        if len(b) == 4:
            seconds = struct.unpack("!L", b)[0]
            nanoseconds = 0
        elif len(b) == 8:
            data64 = struct.unpack("!Q", b)[0]
            seconds = data64 & 0x00000003FFFFFFFF
            nanoseconds = data64 >> 34
        elif len(b) == 12:
            nanoseconds, seconds = struct.unpack("!Iq", b)
        else:
            raise ValueError(
                "Timestamp type can only be created from 32, 64, or 96-bit byte objects"
            )
        return Timestamp(seconds, nanoseconds)

    def to_bytes(self):
        """Pack this Timestamp object into bytes.

        Used for pure-Python msgpack packing.

        :returns data: Payload for EXT message with code -1 (timestamp type)
        :rtype: bytes
        """
        if (self.seconds >> 34) == 0:  # seconds is non-negative and fits in 34 bits
            data64 = self.nanoseconds << 34 | self.seconds
            if data64 & 0xFFFFFFFF00000000 == 0:
                # nanoseconds is zero and seconds < 2**32, so timestamp 32
                data = struct.pack("!L", data64)
            else:
                # timestamp 64
                data = struct.pack("!Q", data64)
        else:
            # timestamp 96
            data = struct.pack("!Iq", self.nanoseconds, self.seconds)
        return data

    @staticmethod
    def from_unix(unix_sec):
        """Create a Timestamp from posix timestamp in seconds.

        :param unix_float: Posix timestamp in seconds.
        :type unix_float: int or float
        """
        seconds = int(unix_sec // 1)
        nanoseconds = int((unix_sec % 1) * 10**9)
        return Timestamp(seconds, nanoseconds)

    def to_unix(self):
        """Get the timestamp as a floating-point value.

        :returns: posix timestamp
        :rtype: float
        """
        return self.seconds + self.nanoseconds / 1e9

    @staticmethod
    def from_unix_nano(unix_ns):
        """Create a Timestamp from posix timestamp in nanoseconds.

        :param int unix_ns: Posix timestamp in nanoseconds.
        :rtype: Timestamp
        """
        return Timestamp(*divmod(unix_ns, 10**9))

    def to_unix_nano(self):
        """Get the timestamp as a unixtime in nanoseconds.

        :returns: posix timestamp in nanoseconds
        :rtype: int
        """
        return self.seconds * 10**9 + self.nanoseconds

    def to_datetime(self):
        """Get the timestamp as a UTC datetime.

        :rtype: `datetime.datetime`
        """
        utc = datetime.timezone.utc
        return datetime.datetime.fromtimestamp(0, utc) + datetime.timedelta(
            seconds=self.seconds, microseconds=self.nanoseconds // 1000
        )

    @staticmethod
    def from_datetime(dt):
        """Create a Timestamp from datetime with tzinfo.

        :rtype: Timestamp
        """
        return Timestamp(seconds=int(dt.timestamp()), nanoseconds=dt.microsecond * 1000)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v135\cast.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: Cast (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

@dataclass
class Sink:
    name: str

    id_: str

    #: Text describing the current session. Present only if there is an active
    #: session on the sink.
    session: typing.Optional[str] = None

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['id'] = self.id_
        if self.session is not None:
            json['session'] = self.session
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            id_=str(json['id']),
            session=str(json['session']) if 'session' in json else None,
        )


def enable(
        presentation_url: typing.Optional[str] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Starts observing for sinks that can be used for tab mirroring, and if set,
    sinks compatible with ``presentationUrl`` as well. When sinks are found, a
    ``sinksUpdated`` event is fired.
    Also starts observing for issue messages. When an issue is added or removed,
    an ``issueUpdated`` event is fired.

    :param presentation_url: *(Optional)*
    '''
    params: T_JSON_DICT = dict()
    if presentation_url is not None:
        params['presentationUrl'] = presentation_url
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.enable',
        'params': params,
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stops observing for sinks and issues.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.disable',
    }
    json = yield cmd_dict


def set_sink_to_use(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Sets a sink to be used when the web page requests the browser to choose a
    sink via Presentation API, Remote Playback API, or Cast SDK.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.setSinkToUse',
        'params': params,
    }
    json = yield cmd_dict


def start_desktop_mirroring(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Starts mirroring the desktop to the sink.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.startDesktopMirroring',
        'params': params,
    }
    json = yield cmd_dict


def start_tab_mirroring(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Starts mirroring the tab to the sink.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.startTabMirroring',
        'params': params,
    }
    json = yield cmd_dict


def stop_casting(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stops the active Cast session on the sink.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.stopCasting',
        'params': params,
    }
    json = yield cmd_dict


@event_class('Cast.sinksUpdated')
@dataclass
class SinksUpdated:
    '''
    This is fired whenever the list of available sinks changes. A sink is a
    device or a software surface that you can cast to.
    '''
    sinks: typing.List[Sink]

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> SinksUpdated:
        return cls(
            sinks=[Sink.from_json(i) for i in json['sinks']]
        )


@event_class('Cast.issueUpdated')
@dataclass
class IssueUpdated:
    '''
    This is fired whenever the outstanding issue/error message changes.
    ``issueMessage`` is empty if there is no issue.
    '''
    issue_message: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> IssueUpdated:
        return cls(
            issue_message=str(json['issueMessage'])
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v136\cast.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: Cast (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

@dataclass
class Sink:
    name: str

    id_: str

    #: Text describing the current session. Present only if there is an active
    #: session on the sink.
    session: typing.Optional[str] = None

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['id'] = self.id_
        if self.session is not None:
            json['session'] = self.session
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            id_=str(json['id']),
            session=str(json['session']) if 'session' in json else None,
        )


def enable(
        presentation_url: typing.Optional[str] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Starts observing for sinks that can be used for tab mirroring, and if set,
    sinks compatible with ``presentationUrl`` as well. When sinks are found, a
    ``sinksUpdated`` event is fired.
    Also starts observing for issue messages. When an issue is added or removed,
    an ``issueUpdated`` event is fired.

    :param presentation_url: *(Optional)*
    '''
    params: T_JSON_DICT = dict()
    if presentation_url is not None:
        params['presentationUrl'] = presentation_url
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.enable',
        'params': params,
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stops observing for sinks and issues.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.disable',
    }
    json = yield cmd_dict


def set_sink_to_use(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Sets a sink to be used when the web page requests the browser to choose a
    sink via Presentation API, Remote Playback API, or Cast SDK.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.setSinkToUse',
        'params': params,
    }
    json = yield cmd_dict


def start_desktop_mirroring(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Starts mirroring the desktop to the sink.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.startDesktopMirroring',
        'params': params,
    }
    json = yield cmd_dict


def start_tab_mirroring(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Starts mirroring the tab to the sink.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.startTabMirroring',
        'params': params,
    }
    json = yield cmd_dict


def stop_casting(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stops the active Cast session on the sink.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.stopCasting',
        'params': params,
    }
    json = yield cmd_dict


@event_class('Cast.sinksUpdated')
@dataclass
class SinksUpdated:
    '''
    This is fired whenever the list of available sinks changes. A sink is a
    device or a software surface that you can cast to.
    '''
    sinks: typing.List[Sink]

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> SinksUpdated:
        return cls(
            sinks=[Sink.from_json(i) for i in json['sinks']]
        )


@event_class('Cast.issueUpdated')
@dataclass
class IssueUpdated:
    '''
    This is fired whenever the outstanding issue/error message changes.
    ``issueMessage`` is empty if there is no issue.
    '''
    issue_message: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> IssueUpdated:
        return cls(
            issue_message=str(json['issueMessage'])
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v137\cast.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: Cast (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

@dataclass
class Sink:
    name: str

    id_: str

    #: Text describing the current session. Present only if there is an active
    #: session on the sink.
    session: typing.Optional[str] = None

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['id'] = self.id_
        if self.session is not None:
            json['session'] = self.session
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            id_=str(json['id']),
            session=str(json['session']) if 'session' in json else None,
        )


def enable(
        presentation_url: typing.Optional[str] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Starts observing for sinks that can be used for tab mirroring, and if set,
    sinks compatible with ``presentationUrl`` as well. When sinks are found, a
    ``sinksUpdated`` event is fired.
    Also starts observing for issue messages. When an issue is added or removed,
    an ``issueUpdated`` event is fired.

    :param presentation_url: *(Optional)*
    '''
    params: T_JSON_DICT = dict()
    if presentation_url is not None:
        params['presentationUrl'] = presentation_url
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.enable',
        'params': params,
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stops observing for sinks and issues.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.disable',
    }
    json = yield cmd_dict


def set_sink_to_use(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Sets a sink to be used when the web page requests the browser to choose a
    sink via Presentation API, Remote Playback API, or Cast SDK.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.setSinkToUse',
        'params': params,
    }
    json = yield cmd_dict


def start_desktop_mirroring(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Starts mirroring the desktop to the sink.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.startDesktopMirroring',
        'params': params,
    }
    json = yield cmd_dict


def start_tab_mirroring(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Starts mirroring the tab to the sink.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.startTabMirroring',
        'params': params,
    }
    json = yield cmd_dict


def stop_casting(
        sink_name: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stops the active Cast session on the sink.

    :param sink_name:
    '''
    params: T_JSON_DICT = dict()
    params['sinkName'] = sink_name
    cmd_dict: T_JSON_DICT = {
        'method': 'Cast.stopCasting',
        'params': params,
    }
    json = yield cmd_dict


@event_class('Cast.sinksUpdated')
@dataclass
class SinksUpdated:
    '''
    This is fired whenever the list of available sinks changes. A sink is a
    device or a software surface that you can cast to.
    '''
    sinks: typing.List[Sink]

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> SinksUpdated:
        return cls(
            sinks=[Sink.from_json(i) for i in json['sinks']]
        )


@event_class('Cast.issueUpdated')
@dataclass
class IssueUpdated:
    '''
    This is fired whenever the outstanding issue/error message changes.
    ``issueMessage`` is empty if there is no issue.
    '''
    issue_message: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> IssueUpdated:
        return cls(
            issue_message=str(json['issueMessage'])
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\orm\__init__.py ===
# orm/__init__.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

"""
Functional constructs for ORM configuration.

See the SQLAlchemy object relational tutorial and mapper configuration
documentation for an overview of how this module is used.

"""

from __future__ import annotations

from typing import Any

from . import exc as exc
from . import mapper as mapperlib
from . import strategy_options as strategy_options
from ._orm_constructors import _mapper_fn as mapper
from ._orm_constructors import aliased as aliased
from ._orm_constructors import backref as backref
from ._orm_constructors import clear_mappers as clear_mappers
from ._orm_constructors import column_property as column_property
from ._orm_constructors import composite as composite
from ._orm_constructors import contains_alias as contains_alias
from ._orm_constructors import create_session as create_session
from ._orm_constructors import deferred as deferred
from ._orm_constructors import dynamic_loader as dynamic_loader
from ._orm_constructors import join as join
from ._orm_constructors import mapped_column as mapped_column
from ._orm_constructors import orm_insert_sentinel as orm_insert_sentinel
from ._orm_constructors import outerjoin as outerjoin
from ._orm_constructors import query_expression as query_expression
from ._orm_constructors import relationship as relationship
from ._orm_constructors import synonym as synonym
from ._orm_constructors import with_loader_criteria as with_loader_criteria
from ._orm_constructors import with_polymorphic as with_polymorphic
from .attributes import AttributeEventToken as AttributeEventToken
from .attributes import InstrumentedAttribute as InstrumentedAttribute
from .attributes import QueryableAttribute as QueryableAttribute
from .base import class_mapper as class_mapper
from .base import DynamicMapped as DynamicMapped
from .base import InspectionAttrExtensionType as InspectionAttrExtensionType
from .base import LoaderCallableStatus as LoaderCallableStatus
from .base import Mapped as Mapped
from .base import NotExtension as NotExtension
from .base import ORMDescriptor as ORMDescriptor
from .base import PassiveFlag as PassiveFlag
from .base import SQLORMExpression as SQLORMExpression
from .base import WriteOnlyMapped as WriteOnlyMapped
from .context import FromStatement as FromStatement
from .context import QueryContext as QueryContext
from .decl_api import add_mapped_attribute as add_mapped_attribute
from .decl_api import as_declarative as as_declarative
from .decl_api import declarative_base as declarative_base
from .decl_api import declarative_mixin as declarative_mixin
from .decl_api import DeclarativeBase as DeclarativeBase
from .decl_api import DeclarativeBaseNoMeta as DeclarativeBaseNoMeta
from .decl_api import DeclarativeMeta as DeclarativeMeta
from .decl_api import declared_attr as declared_attr
from .decl_api import has_inherited_table as has_inherited_table
from .decl_api import MappedAsDataclass as MappedAsDataclass
from .decl_api import registry as registry
from .decl_api import synonym_for as synonym_for
from .decl_base import MappedClassProtocol as MappedClassProtocol
from .descriptor_props import Composite as Composite
from .descriptor_props import CompositeProperty as CompositeProperty
from .descriptor_props import Synonym as Synonym
from .descriptor_props import SynonymProperty as SynonymProperty
from .dynamic import AppenderQuery as AppenderQuery
from .events import AttributeEvents as AttributeEvents
from .events import InstanceEvents as InstanceEvents
from .events import InstrumentationEvents as InstrumentationEvents
from .events import MapperEvents as MapperEvents
from .events import QueryEvents as QueryEvents
from .events import SessionEvents as SessionEvents
from .identity import IdentityMap as IdentityMap
from .instrumentation import ClassManager as ClassManager
from .interfaces import EXT_CONTINUE as EXT_CONTINUE
from .interfaces import EXT_SKIP as EXT_SKIP
from .interfaces import EXT_STOP as EXT_STOP
from .interfaces import InspectionAttr as InspectionAttr
from .interfaces import InspectionAttrInfo as InspectionAttrInfo
from .interfaces import MANYTOMANY as MANYTOMANY
from .interfaces import MANYTOONE as MANYTOONE
from .interfaces import MapperProperty as MapperProperty
from .interfaces import NO_KEY as NO_KEY
from .interfaces import NO_VALUE as NO_VALUE
from .interfaces import ONETOMANY as ONETOMANY
from .interfaces import PropComparator as PropComparator
from .interfaces import RelationshipDirection as RelationshipDirection
from .interfaces import UserDefinedOption as UserDefinedOption
from .loading import merge_frozen_result as merge_frozen_result
from .loading import merge_result as merge_result
from .mapped_collection import attribute_keyed_dict as attribute_keyed_dict
from .mapped_collection import (
    attribute_mapped_collection as attribute_mapped_collection,
)
from .mapped_collection import column_keyed_dict as column_keyed_dict
from .mapped_collection import (
    column_mapped_collection as column_mapped_collection,
)
from .mapped_collection import keyfunc_mapping as keyfunc_mapping
from .mapped_collection import KeyFuncDict as KeyFuncDict
from .mapped_collection import mapped_collection as mapped_collection
from .mapped_collection import MappedCollection as MappedCollection
from .mapper import configure_mappers as configure_mappers
from .mapper import Mapper as Mapper
from .mapper import reconstructor as reconstructor
from .mapper import validates as validates
from .properties import ColumnProperty as ColumnProperty
from .properties import MappedColumn as MappedColumn
from .properties import MappedSQLExpression as MappedSQLExpression
from .query import AliasOption as AliasOption
from .query import Query as Query
from .relationships import foreign as foreign
from .relationships import Relationship as Relationship
from .relationships import RelationshipProperty as RelationshipProperty
from .relationships import remote as remote
from .scoping import QueryPropertyDescriptor as QueryPropertyDescriptor
from .scoping import scoped_session as scoped_session
from .session import close_all_sessions as close_all_sessions
from .session import make_transient as make_transient
from .session import make_transient_to_detached as make_transient_to_detached
from .session import object_session as object_session
from .session import ORMExecuteState as ORMExecuteState
from .session import Session as Session
from .session import sessionmaker as sessionmaker
from .session import SessionTransaction as SessionTransaction
from .session import SessionTransactionOrigin as SessionTransactionOrigin
from .state import AttributeState as AttributeState
from .state import InstanceState as InstanceState
from .strategy_options import contains_eager as contains_eager
from .strategy_options import defaultload as defaultload
from .strategy_options import defer as defer
from .strategy_options import immediateload as immediateload
from .strategy_options import joinedload as joinedload
from .strategy_options import lazyload as lazyload
from .strategy_options import Load as Load
from .strategy_options import load_only as load_only
from .strategy_options import noload as noload
from .strategy_options import raiseload as raiseload
from .strategy_options import selectin_polymorphic as selectin_polymorphic
from .strategy_options import selectinload as selectinload
from .strategy_options import subqueryload as subqueryload
from .strategy_options import undefer as undefer
from .strategy_options import undefer_group as undefer_group
from .strategy_options import with_expression as with_expression
from .unitofwork import UOWTransaction as UOWTransaction
from .util import Bundle as Bundle
from .util import CascadeOptions as CascadeOptions
from .util import LoaderCriteriaOption as LoaderCriteriaOption
from .util import object_mapper as object_mapper
from .util import polymorphic_union as polymorphic_union
from .util import was_deleted as was_deleted
from .util import with_parent as with_parent
from .writeonly import WriteOnlyCollection as WriteOnlyCollection
from .. import util as _sa_util


def __go(lcls: Any) -> None:
    _sa_util.preloaded.import_prefix("sqlalchemy.orm")
    _sa_util.preloaded.import_prefix("sqlalchemy.ext")


__go(locals())

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\liealgebras\type_b.py ===
from .cartan_type import Standard_Cartan
from sympy.core.backend import eye

class TypeB(Standard_Cartan):

    def __new__(cls, n):
        if n < 2:
            raise ValueError("n cannot be less than 2")
        return Standard_Cartan.__new__(cls, "B", n)

    def dimension(self):
        """Dimension of the vector space V underlying the Lie algebra

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType("B3")
        >>> c.dimension()
        3
        """

        return self.n

    def basic_root(self, i, j):
        """
        This is a method just to generate roots
        with a 1 iin the ith position and a -1
        in the jth position.

        """
        root = [0]*self.n
        root[i] = 1
        root[j] = -1
        return root

    def simple_root(self, i):
        """
        Every lie algebra has a unique root system.
        Given a root system Q, there is a subset of the
        roots such that an element of Q is called a
        simple root if it cannot be written as the sum
        of two elements in Q.  If we let D denote the
        set of simple roots, then it is clear that every
        element of Q can be written as a linear combination
        of elements of D with all coefficients non-negative.

        In B_n the first n-1 simple roots are the same as the
        roots in A_(n-1) (a 1 in the ith position, a -1 in
        the (i+1)th position, and zeroes elsewhere).  The n-th
        simple root is the root with a 1 in the nth position
        and zeroes elsewhere.

        This method returns the ith simple root for the B series.

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType("B3")
        >>> c.simple_root(2)
        [0, 1, -1]

        """
        n = self.n
        if i < n:
            return self.basic_root(i-1, i)
        else:
            root = [0]*self.n
            root[n-1] = 1
            return root

    def positive_roots(self):
        """
        This method generates all the positive roots of
        A_n.  This is half of all of the roots of B_n;
        by multiplying all the positive roots by -1 we
        get the negative roots.

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType("A3")
        >>> c.positive_roots()
        {1: [1, -1, 0, 0], 2: [1, 0, -1, 0], 3: [1, 0, 0, -1], 4: [0, 1, -1, 0],
                5: [0, 1, 0, -1], 6: [0, 0, 1, -1]}
        """

        n = self.n
        posroots = {}
        k = 0
        for i in range(0, n-1):
            for j in range(i+1, n):
                k += 1
                posroots[k] = self.basic_root(i, j)
                k += 1
                root = self.basic_root(i, j)
                root[j] = 1
                posroots[k] = root

        for i in range(0, n):
            k += 1
            root = [0]*n
            root[i] = 1
            posroots[k] = root

        return posroots

    def roots(self):
        """
        Returns the total number of roots for B_n"
        """

        n = self.n
        return 2*(n**2)

    def cartan_matrix(self):
        """
        Returns the Cartan matrix for B_n.
        The Cartan matrix matrix for a Lie algebra is
        generated by assigning an ordering to the simple
        roots, (alpha[1], ...., alpha[l]).  Then the ijth
        entry of the Cartan matrix is (<alpha[i],alpha[j]>).

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType('B4')
        >>> c.cartan_matrix()
        Matrix([
        [ 2, -1,  0,  0],
        [-1,  2, -1,  0],
        [ 0, -1,  2, -2],
        [ 0,  0, -1,  2]])

        """

        n = self.n
        m = 2* eye(n)
        for i in range(1, n - 1):
            m[i, i+1] = -1
            m[i, i-1] = -1
        m[0, 1] = -1
        m[n-2, n-1] = -2
        m[n-1, n-2] = -1
        return m

    def basis(self):
        """
        Returns the number of independent generators of B_n
        """

        n = self.n
        return (n**2 - n)/2

    def lie_algebra(self):
        """
        Returns the Lie algebra associated with B_n
        """

        n = self.n
        return "so(" + str(2*n) + ")"

    def dynkin_diagram(self):
        n = self.n
        diag = "---".join("0" for i in range(1, n)) + "=>=0\n"
        diag += "   ".join(str(i) for i in range(1, n+1))
        return diag

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\test_highlevel_ssl_helpers.py ===
from __future__ import annotations

from functools import partial
from typing import TYPE_CHECKING, NoReturn, cast

import attrs
import pytest

import trio
import trio.testing
from trio.socket import AF_INET, IPPROTO_TCP, SOCK_STREAM

from .._highlevel_ssl_helpers import (
    open_ssl_over_tcp_listeners,
    open_ssl_over_tcp_stream,
    serve_ssl_over_tcp,
)

# using noqa because linters don't understand how pytest fixtures work.
from .test_ssl import SERVER_CTX, client_ctx  # noqa: F401

if TYPE_CHECKING:
    from socket import AddressFamily, SocketKind
    from ssl import SSLContext

    from trio.abc import Stream

    from .._highlevel_socket import SocketListener
    from .._ssl import SSLListener


async def echo_handler(stream: Stream) -> None:
    async with stream:
        try:
            while True:
                data = await stream.receive_some(10000)
                if not data:
                    break
                await stream.send_all(data)
        except trio.BrokenResourceError:
            pass


# Resolver that always returns the given sockaddr, no matter what host/port
# you ask for.
@attrs.define(slots=False)
class FakeHostnameResolver(trio.abc.HostnameResolver):
    sockaddr: tuple[str, int] | tuple[str, int, int, int] | tuple[int, bytes]

    async def getaddrinfo(
        self,
        host: bytes | None,
        port: bytes | str | int | None,
        family: int = 0,
        type: int = 0,
        proto: int = 0,
        flags: int = 0,
    ) -> list[
        tuple[
            AddressFamily,
            SocketKind,
            int,
            str,
            tuple[str, int] | tuple[str, int, int, int] | tuple[int, bytes],
        ]
    ]:
        return [(AF_INET, SOCK_STREAM, IPPROTO_TCP, "", self.sockaddr)]

    async def getnameinfo(
        self,
        sockaddr: tuple[str, int] | tuple[str, int, int, int],
        flags: int,
    ) -> NoReturn:  # pragma: no cover
        raise NotImplementedError


# This uses serve_ssl_over_tcp, which uses open_ssl_over_tcp_listeners...
# using noqa because linters don't understand how pytest fixtures work.
async def test_open_ssl_over_tcp_stream_and_everything_else(
    client_ctx: SSLContext,  # noqa: F811 # linters doesn't understand fixture
) -> None:
    async with trio.open_nursery() as nursery:
        # TODO: this function wraps an SSLListener around a SocketListener, this is illegal
        # according to current type hints, and probably for good reason. But there should
        # maybe be a different wrapper class/function that could be used instead?
        value = await nursery.start(
            partial(
                serve_ssl_over_tcp,
                echo_handler,
                0,
                SERVER_CTX,
                host="127.0.0.1",
            ),
        )
        assert isinstance(value, list)
        res = cast("list[SSLListener[SocketListener]]", value)  # type: ignore[type-var]
        (listener,) = res
        async with listener:
            # listener.transport_listener is of type Listener[Stream]
            tp_listener: SocketListener = listener.transport_listener  # type: ignore[assignment]

            sockaddr = tp_listener.socket.getsockname()
            hostname_resolver = FakeHostnameResolver(sockaddr)
            trio.socket.set_custom_hostname_resolver(hostname_resolver)

            # We don't have the right trust set up
            # (checks that ssl_context=None is doing some validation)
            stream = await open_ssl_over_tcp_stream("trio-test-1.example.org", 80)
            async with stream:
                with pytest.raises(trio.BrokenResourceError):
                    await stream.do_handshake()

            # We have the trust but not the hostname
            # (checks custom ssl_context + hostname checking)
            stream = await open_ssl_over_tcp_stream(
                "xyzzy.example.org",
                80,
                ssl_context=client_ctx,
            )
            async with stream:
                with pytest.raises(trio.BrokenResourceError):
                    await stream.do_handshake()

            # This one should work!
            stream = await open_ssl_over_tcp_stream(
                "trio-test-1.example.org",
                80,
                ssl_context=client_ctx,
            )
            async with stream:
                assert isinstance(stream, trio.SSLStream)
                assert stream.server_hostname == "trio-test-1.example.org"
                await stream.send_all(b"x")
                assert await stream.receive_some(1) == b"x"

            # Check https_compatible settings are being passed through
            assert not stream._https_compatible
            stream = await open_ssl_over_tcp_stream(
                "trio-test-1.example.org",
                80,
                ssl_context=client_ctx,
                https_compatible=True,
                # also, smoke test happy_eyeballs_delay
                happy_eyeballs_delay=1,
            )
            async with stream:
                assert stream._https_compatible

            # Stop the echo server
            nursery.cancel_scope.cancel()


async def test_open_ssl_over_tcp_listeners() -> None:
    (listener,) = await open_ssl_over_tcp_listeners(0, SERVER_CTX, host="127.0.0.1")
    async with listener:
        assert isinstance(listener, trio.SSLListener)
        tl = listener.transport_listener
        assert isinstance(tl, trio.SocketListener)
        assert tl.socket.getsockname()[0] == "127.0.0.1"

        assert not listener._https_compatible

    (listener,) = await open_ssl_over_tcp_listeners(
        0,
        SERVER_CTX,
        host="127.0.0.1",
        https_compatible=True,
    )
    async with listener:
        assert listener._https_compatible

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\werkzeug\sansio\http.py ===
from __future__ import annotations

import re
import typing as t
from datetime import datetime

from .._internal import _dt_as_utc
from ..http import generate_etag
from ..http import parse_date
from ..http import parse_etags
from ..http import parse_if_range_header
from ..http import unquote_etag

_etag_re = re.compile(r'([Ww]/)?(?:"(.*?)"|(.*?))(?:\s*,\s*|$)')


def is_resource_modified(
    http_range: str | None = None,
    http_if_range: str | None = None,
    http_if_modified_since: str | None = None,
    http_if_none_match: str | None = None,
    http_if_match: str | None = None,
    etag: str | None = None,
    data: bytes | None = None,
    last_modified: datetime | str | None = None,
    ignore_if_range: bool = True,
) -> bool:
    """Convenience method for conditional requests.
    :param http_range: Range HTTP header
    :param http_if_range: If-Range HTTP header
    :param http_if_modified_since: If-Modified-Since HTTP header
    :param http_if_none_match: If-None-Match HTTP header
    :param http_if_match: If-Match HTTP header
    :param etag: the etag for the response for comparison.
    :param data: or alternatively the data of the response to automatically
                 generate an etag using :func:`generate_etag`.
    :param last_modified: an optional date of the last modification.
    :param ignore_if_range: If `False`, `If-Range` header will be taken into
                            account.
    :return: `True` if the resource was modified, otherwise `False`.

    .. versionadded:: 2.2
    """
    if etag is None and data is not None:
        etag = generate_etag(data)
    elif data is not None:
        raise TypeError("both data and etag given")

    unmodified = False
    if isinstance(last_modified, str):
        last_modified = parse_date(last_modified)

    # HTTP doesn't use microsecond, remove it to avoid false positive
    # comparisons. Mark naive datetimes as UTC.
    if last_modified is not None:
        last_modified = _dt_as_utc(last_modified.replace(microsecond=0))

    if_range = None
    if not ignore_if_range and http_range is not None:
        # https://tools.ietf.org/html/rfc7233#section-3.2
        # A server MUST ignore an If-Range header field received in a request
        # that does not contain a Range header field.
        if_range = parse_if_range_header(http_if_range)

    if if_range is not None and if_range.date is not None:
        modified_since: datetime | None = if_range.date
    else:
        modified_since = parse_date(http_if_modified_since)

    if modified_since and last_modified and last_modified <= modified_since:
        unmodified = True

    if etag:
        etag, _ = unquote_etag(etag)

        if if_range is not None and if_range.etag is not None:
            unmodified = parse_etags(if_range.etag).contains(etag)
        else:
            if_none_match = parse_etags(http_if_none_match)
            if if_none_match:
                # https://tools.ietf.org/html/rfc7232#section-3.2
                # "A recipient MUST use the weak comparison function when comparing
                # entity-tags for If-None-Match"
                unmodified = if_none_match.contains_weak(etag)

            # https://tools.ietf.org/html/rfc7232#section-3.1
            # "Origin server MUST use the strong comparison function when
            # comparing entity-tags for If-Match"
            if_match = parse_etags(http_if_match)
            if if_match:
                unmodified = not if_match.is_strong(etag)

    return not unmodified


_cookie_re = re.compile(
    r"""
    ([^=;]*)
    (?:\s*=\s*
      (
        "(?:[^\\"]|\\.)*"
      |
        .*?
      )
    )?
    \s*;\s*
    """,
    flags=re.ASCII | re.VERBOSE,
)
_cookie_unslash_re = re.compile(rb"\\([0-3][0-7]{2}|.)")


def _cookie_unslash_replace(m: t.Match[bytes]) -> bytes:
    v = m.group(1)

    if len(v) == 1:
        return v

    return int(v, 8).to_bytes(1, "big")


def parse_cookie(
    cookie: str | None = None,
    cls: type[ds.MultiDict[str, str]] | None = None,
) -> ds.MultiDict[str, str]:
    """Parse a cookie from a string.

    The same key can be provided multiple times, the values are stored
    in-order. The default :class:`MultiDict` will have the first value
    first, and all values can be retrieved with
    :meth:`MultiDict.getlist`.

    :param cookie: The cookie header as a string.
    :param cls: A dict-like class to store the parsed cookies in.
        Defaults to :class:`MultiDict`.

    .. versionchanged:: 3.0
        Passing bytes, and the ``charset`` and ``errors`` parameters, were removed.

    .. versionadded:: 2.2
    """
    if cls is None:
        cls = t.cast("type[ds.MultiDict[str, str]]", ds.MultiDict)

    if not cookie:
        return cls()

    cookie = f"{cookie};"
    out = []

    for ck, cv in _cookie_re.findall(cookie):
        ck = ck.strip()
        cv = cv.strip()

        if not ck:
            continue

        if len(cv) >= 2 and cv[0] == cv[-1] == '"':
            # Work with bytes here, since a UTF-8 character could be multiple bytes.
            cv = _cookie_unslash_re.sub(
                _cookie_unslash_replace, cv[1:-1].encode()
            ).decode(errors="replace")

        out.append((ck, cv))

    return cls(out)


# circular dependencies
from .. import datastructures as ds

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\wtforms\fields\datetime.py ===
import datetime

from wtforms import widgets
from wtforms.fields.core import Field
from wtforms.utils import clean_datetime_format_for_strptime

__all__ = (
    "DateTimeField",
    "DateField",
    "TimeField",
    "MonthField",
    "DateTimeLocalField",
    "WeekField",
)


class DateTimeField(Field):
    """
    A text field which stores a :class:`datetime.datetime` matching one or
    several formats. If ``format`` is a list, any input value matching any
    format will be accepted, and the first format in the list will be used
    to produce HTML values.
    """

    widget = widgets.DateTimeInput()

    def __init__(
        self, label=None, validators=None, format="%Y-%m-%d %H:%M:%S", **kwargs
    ):
        super().__init__(label, validators, **kwargs)
        self.format = format if isinstance(format, list) else [format]
        self.strptime_format = clean_datetime_format_for_strptime(self.format)

    def _value(self):
        if self.raw_data:
            return " ".join(self.raw_data)
        format = self.format[0]
        return self.data and self.data.strftime(format) or ""

    def process_formdata(self, valuelist):
        if not valuelist:
            return

        date_str = " ".join(valuelist)
        for format in self.strptime_format:
            try:
                self.data = datetime.datetime.strptime(date_str, format)
                return
            except ValueError:
                self.data = None

        raise ValueError(self.gettext("Not a valid datetime value."))


class DateField(DateTimeField):
    """
    Same as :class:`~wtforms.fields.DateTimeField`, except stores a
    :class:`datetime.date`.
    """

    widget = widgets.DateInput()

    def __init__(self, label=None, validators=None, format="%Y-%m-%d", **kwargs):
        super().__init__(label, validators, format, **kwargs)

    def process_formdata(self, valuelist):
        if not valuelist:
            return

        date_str = " ".join(valuelist)
        for format in self.strptime_format:
            try:
                self.data = datetime.datetime.strptime(date_str, format).date()
                return
            except ValueError:
                self.data = None

        raise ValueError(self.gettext("Not a valid date value."))


class TimeField(DateTimeField):
    """
    Same as :class:`~wtforms.fields.DateTimeField`, except stores a
    :class:`datetime.time`.
    """

    widget = widgets.TimeInput()

    def __init__(self, label=None, validators=None, format="%H:%M", **kwargs):
        super().__init__(label, validators, format, **kwargs)

    def process_formdata(self, valuelist):
        if not valuelist:
            return

        time_str = " ".join(valuelist)
        for format in self.strptime_format:
            try:
                self.data = datetime.datetime.strptime(time_str, format).time()
                return
            except ValueError:
                self.data = None

        raise ValueError(self.gettext("Not a valid time value."))


class MonthField(DateField):
    """
    Same as :class:`~wtforms.fields.DateField`, except represents a month,
    stores a :class:`datetime.date` with `day = 1`.
    """

    widget = widgets.MonthInput()

    def __init__(self, label=None, validators=None, format="%Y-%m", **kwargs):
        super().__init__(label, validators, format, **kwargs)


class WeekField(DateField):
    """
    Same as :class:`~wtforms.fields.DateField`, except represents a week,
    stores a :class:`datetime.date` of the monday of the given week.
    """

    widget = widgets.WeekInput()

    def __init__(self, label=None, validators=None, format="%Y-W%W", **kwargs):
        super().__init__(label, validators, format, **kwargs)

    def process_formdata(self, valuelist):
        if not valuelist:
            return

        time_str = " ".join(valuelist)
        for format in self.strptime_format:
            try:
                if "%w" not in format:
                    # The '%w' week starting day is needed. This defaults it to monday
                    # like ISO 8601 indicates.
                    self.data = datetime.datetime.strptime(
                        f"{time_str}-1", f"{format}-%w"
                    ).date()
                else:
                    self.data = datetime.datetime.strptime(time_str, format).date()
                return
            except ValueError:
                self.data = None

        raise ValueError(self.gettext("Not a valid week value."))


class DateTimeLocalField(DateTimeField):
    """
    Same as :class:`~wtforms.fields.DateTimeField`, but represents an
    ``<input type="datetime-local">``.
    """

    widget = widgets.DateTimeLocalInput()

    def __init__(self, *args, **kwargs):
        kwargs.setdefault(
            "format",
            [
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%Y-%m-%dT%H:%M",
            ],
        )
        super().__init__(*args, **kwargs)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\tools\offline_tuning.py ===
# Copyright (c) Microsoft Corporation. All rights reserved.
# Licensed under the MIT License.

import argparse
import copy
import json
import sys
from collections import OrderedDict
from pprint import pprint
from typing import Any

import onnx

TuningResults = dict[str, Any]

_TUNING_RESULTS_KEY = "tuning_results"


def _find_tuning_results_in_props(metadata_props):
    for idx, prop in enumerate(metadata_props):
        if prop.key == _TUNING_RESULTS_KEY:
            return idx
    return -1


def extract(model: onnx.ModelProto):
    idx = _find_tuning_results_in_props(model.metadata_props)
    if idx < 0:
        return None

    tuning_results_prop = model.metadata_props[idx]
    return json.loads(tuning_results_prop.value)


def embed(model: onnx.ModelProto, tuning_results: list[TuningResults], overwrite=False):
    idx = _find_tuning_results_in_props(model.metadata_props)
    assert overwrite or idx <= 0, "the supplied onnx file already have tuning results embedded!"

    if idx >= 0:
        model.metadata_props.pop(idx)

    entry = model.metadata_props.add()
    entry.key = _TUNING_RESULTS_KEY
    entry.value = json.dumps(tuning_results)
    return model


class Merger:
    class EpAndValidators:
        def __init__(self, ep: str, validators: dict[str, str]):
            self.ep = ep
            self.validators = copy.deepcopy(validators)
            self.key = (ep, tuple(sorted(validators.items())))

        def __hash__(self):
            return hash(self.key)

        def __eq__(self, other):
            return self.ep == other.ep and self.key == other.key

    def __init__(self):
        self.ev_to_results = OrderedDict()

    def merge(self, tuning_results: list[TuningResults]):
        for trs in tuning_results:
            self._merge_one(trs)

    def get_merged(self):
        tuning_results = []
        for ev, flat_results in self.ev_to_results.items():
            results = {}
            trs = {
                "ep": ev.ep,
                "validators": ev.validators,
                "results": results,
            }
            for (op_sig, params_sig), kernel_id in flat_results.items():
                kernel_map = results.setdefault(op_sig, {})
                kernel_map[params_sig] = kernel_id
            tuning_results.append(trs)
        return tuning_results

    def _merge_one(self, trs: TuningResults):
        ev = Merger.EpAndValidators(trs["ep"], trs["validators"])
        flat_results = self.ev_to_results.setdefault(ev, {})
        for op_sig, kernel_map in trs["results"].items():
            for params_sig, kernel_id in kernel_map.items():
                if (op_sig, params_sig) not in flat_results:
                    flat_results[(op_sig, params_sig)] = kernel_id


def parse_args():
    parser = argparse.ArgumentParser()
    sub_parsers = parser.add_subparsers(help="Command to execute", dest="cmd")

    extract_parser = sub_parsers.add_parser("extract", help="Extract embedded tuning results from an onnx file.")
    extract_parser.add_argument("input_onnx")
    extract_parser.add_argument("output_json")

    embed_parser = sub_parsers.add_parser("embed", help="Embed the tuning results into an onnx file.")
    embed_parser.add_argument("--force", "-f", action="store_true", help="Overwrite the tuning results if it existed.")
    embed_parser.add_argument("output_onnx", help="Path of the output onnx file.")
    embed_parser.add_argument("input_onnx", help="Path of the input onnx file.")
    embed_parser.add_argument("input_json", nargs="+", help="Path(s) of the tuning results file(s) to be embedded.")

    merge_parser = sub_parsers.add_parser("merge", help="Merge multiple tuning results files as a single one.")
    merge_parser.add_argument("output_json", help="Path of the output tuning results file.")
    merge_parser.add_argument("input_json", nargs="+", help="Paths of the tuning results files to be merged.")

    pprint_parser = sub_parsers.add_parser("pprint", help="Pretty print the tuning results.")
    pprint_parser.add_argument("json_or_onnx", help="A tuning results json file or an onnx file.")

    args = parser.parse_args()
    if len(vars(args)) == 0:
        parser.print_help()
        exit(-1)
    return args


def main():
    args = parse_args()
    if args.cmd == "extract":
        tuning_results = extract(onnx.load_model(args.input_onnx))
        if tuning_results is None:
            sys.stderr.write(f"{args.input_onnx} does not have tuning results embedded!\n")
            sys.exit(-1)
        json.dump(tuning_results, open(args.output_json, "w"))  # noqa: SIM115
    elif args.cmd == "embed":
        model = onnx.load_model(args.input_onnx)
        merger = Merger()
        for tuning_results in [json.load(open(f)) for f in args.input_json]:  # noqa: SIM115
            merger.merge(tuning_results)
        model = embed(model, merger.get_merged(), args.force)
        onnx.save_model(model, args.output_onnx)
    elif args.cmd == "merge":
        merger = Merger()
        for tuning_results in [json.load(open(f)) for f in args.input_json]:  # noqa: SIM115
            merger.merge(tuning_results)
        json.dump(merger.get_merged(), open(args.output_json, "w"))  # noqa: SIM115
    elif args.cmd == "pprint":
        tuning_results = None
        try:  # noqa: SIM105
            tuning_results = json.load(open(args.json_or_onnx))  # noqa: SIM115
        except Exception:
            # it might be an onnx file otherwise, try it latter
            pass

        if tuning_results is None:
            try:
                model = onnx.load_model(args.json_or_onnx)
                tuning_results = extract(model)
                if tuning_results is None:
                    sys.stderr.write(f"{args.input_onnx} does not have tuning results embedded!\n")
                    sys.exit(-1)
            except Exception:
                pass

        if tuning_results is None:
            sys.stderr.write(f"{args.json_or_onnx} is not a valid tuning results file or onnx file!")
            sys.exit(-1)

        pprint(tuning_results)
    else:
        # invalid choice will be handled by the parser
        pass


if __name__ == "__main__":
    main()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\window\common.py ===
"""Common utility functions for rolling operations"""
from __future__ import annotations

from collections import defaultdict
from typing import cast

import numpy as np

from pandas.core.dtypes.generic import (
    ABCDataFrame,
    ABCSeries,
)

from pandas.core.indexes.api import MultiIndex


def flex_binary_moment(arg1, arg2, f, pairwise: bool = False):
    if isinstance(arg1, ABCSeries) and isinstance(arg2, ABCSeries):
        X, Y = prep_binary(arg1, arg2)
        return f(X, Y)

    elif isinstance(arg1, ABCDataFrame):
        from pandas import DataFrame

        def dataframe_from_int_dict(data, frame_template) -> DataFrame:
            result = DataFrame(data, index=frame_template.index)
            if len(result.columns) > 0:
                result.columns = frame_template.columns[result.columns]
            else:
                result.columns = frame_template.columns.copy()
            return result

        results = {}
        if isinstance(arg2, ABCDataFrame):
            if pairwise is False:
                if arg1 is arg2:
                    # special case in order to handle duplicate column names
                    for i in range(len(arg1.columns)):
                        results[i] = f(arg1.iloc[:, i], arg2.iloc[:, i])
                    return dataframe_from_int_dict(results, arg1)
                else:
                    if not arg1.columns.is_unique:
                        raise ValueError("'arg1' columns are not unique")
                    if not arg2.columns.is_unique:
                        raise ValueError("'arg2' columns are not unique")
                    X, Y = arg1.align(arg2, join="outer")
                    X, Y = prep_binary(X, Y)
                    res_columns = arg1.columns.union(arg2.columns)
                    for col in res_columns:
                        if col in X and col in Y:
                            results[col] = f(X[col], Y[col])
                    return DataFrame(results, index=X.index, columns=res_columns)
            elif pairwise is True:
                results = defaultdict(dict)
                for i in range(len(arg1.columns)):
                    for j in range(len(arg2.columns)):
                        if j < i and arg2 is arg1:
                            # Symmetric case
                            results[i][j] = results[j][i]
                        else:
                            results[i][j] = f(
                                *prep_binary(arg1.iloc[:, i], arg2.iloc[:, j])
                            )

                from pandas import concat

                result_index = arg1.index.union(arg2.index)
                if len(result_index):
                    # construct result frame
                    result = concat(
                        [
                            concat(
                                [results[i][j] for j in range(len(arg2.columns))],
                                ignore_index=True,
                            )
                            for i in range(len(arg1.columns))
                        ],
                        ignore_index=True,
                        axis=1,
                    )
                    result.columns = arg1.columns

                    # set the index and reorder
                    if arg2.columns.nlevels > 1:
                        # mypy needs to know columns is a MultiIndex, Index doesn't
                        # have levels attribute
                        arg2.columns = cast(MultiIndex, arg2.columns)
                        # GH 21157: Equivalent to MultiIndex.from_product(
                        #  [result_index], <unique combinations of arg2.columns.levels>,
                        # )
                        # A normal MultiIndex.from_product will produce too many
                        # combinations.
                        result_level = np.tile(
                            result_index, len(result) // len(result_index)
                        )
                        arg2_levels = (
                            np.repeat(
                                arg2.columns.get_level_values(i),
                                len(result) // len(arg2.columns),
                            )
                            for i in range(arg2.columns.nlevels)
                        )
                        result_names = list(arg2.columns.names) + [result_index.name]
                        result.index = MultiIndex.from_arrays(
                            [*arg2_levels, result_level], names=result_names
                        )
                        # GH 34440
                        num_levels = len(result.index.levels)
                        new_order = [num_levels - 1] + list(range(num_levels - 1))
                        result = result.reorder_levels(new_order).sort_index()
                    else:
                        result.index = MultiIndex.from_product(
                            [range(len(arg2.columns)), range(len(result_index))]
                        )
                        result = result.swaplevel(1, 0).sort_index()
                        result.index = MultiIndex.from_product(
                            [result_index] + [arg2.columns]
                        )
                else:
                    # empty result
                    result = DataFrame(
                        index=MultiIndex(
                            levels=[arg1.index, arg2.columns], codes=[[], []]
                        ),
                        columns=arg2.columns,
                        dtype="float64",
                    )

                # reset our index names to arg1 names
                # reset our column names to arg2 names
                # careful not to mutate the original names
                result.columns = result.columns.set_names(arg1.columns.names)
                result.index = result.index.set_names(
                    result_index.names + arg2.columns.names
                )

                return result
        else:
            results = {
                i: f(*prep_binary(arg1.iloc[:, i], arg2))
                for i in range(len(arg1.columns))
            }
            return dataframe_from_int_dict(results, arg1)

    else:
        return flex_binary_moment(arg2, arg1, f)


def zsqrt(x):
    with np.errstate(all="ignore"):
        result = np.sqrt(x)
        mask = x < 0

    if isinstance(x, ABCDataFrame):
        if mask._values.any():
            result[mask] = 0
    else:
        if mask.any():
            result[mask] = 0

    return result


def prep_binary(arg1, arg2):
    # mask out values, this also makes a common index...
    X = arg1 + 0 * arg2
    Y = arg2 + 0 * arg1

    return X, Y

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\liealgebras\type_c.py ===
from .cartan_type import Standard_Cartan
from sympy.core.backend import eye

class TypeC(Standard_Cartan):

    def __new__(cls, n):
        if n < 3:
            raise ValueError("n cannot be less than 3")
        return Standard_Cartan.__new__(cls, "C", n)


    def dimension(self):
        """Dimension of the vector space V underlying the Lie algebra

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType("C3")
        >>> c.dimension()
        3
        """
        n = self.n
        return n

    def basic_root(self, i, j):
        """Generate roots with 1 in ith position and a -1 in jth position
        """
        n = self.n
        root = [0]*n
        root[i] = 1
        root[j] = -1
        return root

    def simple_root(self, i):
        """The ith simple root for the C series

        Every lie algebra has a unique root system.
        Given a root system Q, there is a subset of the
        roots such that an element of Q is called a
        simple root if it cannot be written as the sum
        of two elements in Q.  If we let D denote the
        set of simple roots, then it is clear that every
        element of Q can be written as a linear combination
        of elements of D with all coefficients non-negative.

        In C_n, the first n-1 simple roots are the same as
        the roots in A_(n-1) (a 1 in the ith position, a -1
        in the (i+1)th position, and zeroes elsewhere).  The
        nth simple root is the root in which there is a 2 in
        the nth position and zeroes elsewhere.

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType("C3")
        >>> c.simple_root(2)
        [0, 1, -1]

        """

        n = self.n
        if i < n:
            return self.basic_root(i-1,i)
        else:
            root = [0]*self.n
            root[n-1] = 2
            return root


    def positive_roots(self):
        """Generates all the positive roots of A_n

        This is half of all of the roots of C_n; by multiplying all the
        positive roots by -1 we get the negative roots.

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType("A3")
        >>> c.positive_roots()
        {1: [1, -1, 0, 0], 2: [1, 0, -1, 0], 3: [1, 0, 0, -1], 4: [0, 1, -1, 0],
                5: [0, 1, 0, -1], 6: [0, 0, 1, -1]}

        """

        n = self.n
        posroots = {}
        k = 0
        for i in range(0, n-1):
            for j in range(i+1, n):
                k += 1
                posroots[k] = self.basic_root(i, j)
                k += 1
                root = self.basic_root(i, j)
                root[j] = 1
                posroots[k] = root

        for i in range(0, n):
            k += 1
            root = [0]*n
            root[i] = 2
            posroots[k] = root

        return posroots

    def roots(self):
        """
        Returns the total number of roots for C_n"
        """

        n = self.n
        return 2*(n**2)

    def cartan_matrix(self):
        """The Cartan matrix for C_n

        The Cartan matrix matrix for a Lie algebra is
        generated by assigning an ordering to the simple
        roots, (alpha[1], ...., alpha[l]).  Then the ijth
        entry of the Cartan matrix is (<alpha[i],alpha[j]>).

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType('C4')
        >>> c.cartan_matrix()
        Matrix([
        [ 2, -1,  0,  0],
        [-1,  2, -1,  0],
        [ 0, -1,  2, -1],
        [ 0,  0, -2,  2]])

        """

        n = self.n
        m = 2 * eye(n)
        for i in range(1, n - 1):
            m[i, i+1] = -1
            m[i, i-1] = -1
        m[0,1] = -1
        m[n-1, n-2] = -2
        return m


    def basis(self):
        """
        Returns the number of independent generators of C_n
        """

        n = self.n
        return n*(2*n + 1)

    def lie_algebra(self):
        """
        Returns the Lie algebra associated with C_n"
        """

        n = self.n
        return "sp(" + str(2*n) + ")"

    def dynkin_diagram(self):
        n = self.n
        diag = "---".join("0" for i in range(1, n)) + "=<=0\n"
        diag += "   ".join(str(i) for i in range(1, n+1))
        return diag

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\compat\_optional.py ===
from __future__ import annotations

import importlib
import sys
from typing import TYPE_CHECKING
import warnings

from pandas.util._exceptions import find_stack_level

from pandas.util.version import Version

if TYPE_CHECKING:
    import types

# Update install.rst & setup.cfg when updating versions!

VERSIONS = {
    "adbc-driver-postgresql": "0.8.0",
    "adbc-driver-sqlite": "0.8.0",
    "bs4": "4.11.2",
    "blosc": "1.21.3",
    "bottleneck": "1.3.6",
    "dataframe-api-compat": "0.1.7",
    "fastparquet": "2022.12.0",
    "fsspec": "2022.11.0",
    "html5lib": "1.1",
    "hypothesis": "6.46.1",
    "gcsfs": "2022.11.0",
    "jinja2": "3.1.2",
    "lxml.etree": "4.9.2",
    "matplotlib": "3.6.3",
    "numba": "0.56.4",
    "numexpr": "2.8.4",
    "odfpy": "1.4.1",
    "openpyxl": "3.1.0",
    "pandas_gbq": "0.19.0",
    "psycopg2": "2.9.6",  # (dt dec pq3 ext lo64)
    "pymysql": "1.0.2",
    "pyarrow": "10.0.1",
    "pyreadstat": "1.2.0",
    "pytest": "7.3.2",
    "python-calamine": "0.1.7",
    "pyxlsb": "1.0.10",
    "s3fs": "2022.11.0",
    "scipy": "1.10.0",
    "sqlalchemy": "2.0.0",
    "tables": "3.8.0",
    "tabulate": "0.9.0",
    "xarray": "2022.12.0",
    "xlrd": "2.0.1",
    "xlsxwriter": "3.0.5",
    "zstandard": "0.19.0",
    "tzdata": "2022.7",
    "qtpy": "2.3.0",
    "pyqt5": "5.15.9",
}

# A mapping from import name to package name (on PyPI) for packages where
# these two names are different.

INSTALL_MAPPING = {
    "bs4": "beautifulsoup4",
    "bottleneck": "Bottleneck",
    "jinja2": "Jinja2",
    "lxml.etree": "lxml",
    "odf": "odfpy",
    "pandas_gbq": "pandas-gbq",
    "python_calamine": "python-calamine",
    "sqlalchemy": "SQLAlchemy",
    "tables": "pytables",
}


def get_version(module: types.ModuleType) -> str:
    version = getattr(module, "__version__", None)

    if version is None:
        raise ImportError(f"Can't determine version for {module.__name__}")
    if module.__name__ == "psycopg2":
        # psycopg2 appends " (dt dec pq3 ext lo64)" to it's version
        version = version.split()[0]
    return version


def import_optional_dependency(
    name: str,
    extra: str = "",
    errors: str = "raise",
    min_version: str | None = None,
):
    """
    Import an optional dependency.

    By default, if a dependency is missing an ImportError with a nice
    message will be raised. If a dependency is present, but too old,
    we raise.

    Parameters
    ----------
    name : str
        The module name.
    extra : str
        Additional text to include in the ImportError message.
    errors : str {'raise', 'warn', 'ignore'}
        What to do when a dependency is not found or its version is too old.

        * raise : Raise an ImportError
        * warn : Only applicable when a module's version is to old.
          Warns that the version is too old and returns None
        * ignore: If the module is not installed, return None, otherwise,
          return the module, even if the version is too old.
          It's expected that users validate the version locally when
          using ``errors="ignore"`` (see. ``io/html.py``)
    min_version : str, default None
        Specify a minimum version that is different from the global pandas
        minimum version required.
    Returns
    -------
    maybe_module : Optional[ModuleType]
        The imported module, when found and the version is correct.
        None is returned when the package is not found and `errors`
        is False, or when the package's version is too old and `errors`
        is ``'warn'`` or ``'ignore'``.
    """
    assert errors in {"warn", "raise", "ignore"}

    package_name = INSTALL_MAPPING.get(name)
    install_name = package_name if package_name is not None else name

    msg = (
        f"Missing optional dependency '{install_name}'. {extra} "
        f"Use pip or conda to install {install_name}."
    )
    try:
        module = importlib.import_module(name)
    except ImportError:
        if errors == "raise":
            raise ImportError(msg)
        return None

    # Handle submodules: if we have submodule, grab parent module from sys.modules
    parent = name.split(".")[0]
    if parent != name:
        install_name = parent
        module_to_get = sys.modules[install_name]
    else:
        module_to_get = module
    minimum_version = min_version if min_version is not None else VERSIONS.get(parent)
    if minimum_version:
        version = get_version(module_to_get)
        if version and Version(version) < Version(minimum_version):
            msg = (
                f"Pandas requires version '{minimum_version}' or newer of '{parent}' "
                f"(version '{version}' currently installed)."
            )
            if errors == "warn":
                warnings.warn(
                    msg,
                    UserWarning,
                    stacklevel=find_stack_level(),
                )
                return None
            elif errors == "raise":
                raise ImportError(msg)
            else:
                return None

    return module

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\tools\times.py ===
from __future__ import annotations

from datetime import (
    datetime,
    time,
)
from typing import TYPE_CHECKING
import warnings

import numpy as np

from pandas._libs.lib import is_list_like
from pandas.util._exceptions import find_stack_level

from pandas.core.dtypes.generic import (
    ABCIndex,
    ABCSeries,
)
from pandas.core.dtypes.missing import notna

if TYPE_CHECKING:
    from pandas._typing import DateTimeErrorChoices


def to_time(
    arg,
    format: str | None = None,
    infer_time_format: bool = False,
    errors: DateTimeErrorChoices = "raise",
):
    """
    Parse time strings to time objects using fixed strptime formats ("%H:%M",
    "%H%M", "%I:%M%p", "%I%M%p", "%H:%M:%S", "%H%M%S", "%I:%M:%S%p",
    "%I%M%S%p")

    Use infer_time_format if all the strings are in the same format to speed
    up conversion.

    Parameters
    ----------
    arg : string in time format, datetime.time, list, tuple, 1-d array,  Series
    format : str, default None
        Format used to convert arg into a time object.  If None, fixed formats
        are used.
    infer_time_format: bool, default False
        Infer the time format based on the first non-NaN element.  If all
        strings are in the same format, this will speed up conversion.
    errors : {'ignore', 'raise', 'coerce'}, default 'raise'
        - If 'raise', then invalid parsing will raise an exception
        - If 'coerce', then invalid parsing will be set as None
        - If 'ignore', then invalid parsing will return the input

    Returns
    -------
    datetime.time
    """
    if errors == "ignore":
        # GH#54467
        warnings.warn(
            "errors='ignore' is deprecated and will raise in a future version. "
            "Use to_time without passing `errors` and catch exceptions "
            "explicitly instead",
            FutureWarning,
            stacklevel=find_stack_level(),
        )

    def _convert_listlike(arg, format):
        if isinstance(arg, (list, tuple)):
            arg = np.array(arg, dtype="O")

        elif getattr(arg, "ndim", 1) > 1:
            raise TypeError(
                "arg must be a string, datetime, list, tuple, 1-d array, or Series"
            )

        arg = np.asarray(arg, dtype="O")

        if infer_time_format and format is None:
            format = _guess_time_format_for_array(arg)

        times: list[time | None] = []
        if format is not None:
            for element in arg:
                try:
                    times.append(datetime.strptime(element, format).time())
                except (ValueError, TypeError) as err:
                    if errors == "raise":
                        msg = (
                            f"Cannot convert {element} to a time with given "
                            f"format {format}"
                        )
                        raise ValueError(msg) from err
                    if errors == "ignore":
                        return arg
                    else:
                        times.append(None)
        else:
            formats = _time_formats[:]
            format_found = False
            for element in arg:
                time_object = None
                try:
                    time_object = time.fromisoformat(element)
                except (ValueError, TypeError):
                    for time_format in formats:
                        try:
                            time_object = datetime.strptime(element, time_format).time()
                            if not format_found:
                                # Put the found format in front
                                fmt = formats.pop(formats.index(time_format))
                                formats.insert(0, fmt)
                                format_found = True
                            break
                        except (ValueError, TypeError):
                            continue

                if time_object is not None:
                    times.append(time_object)
                elif errors == "raise":
                    raise ValueError(f"Cannot convert arg {arg} to a time")
                elif errors == "ignore":
                    return arg
                else:
                    times.append(None)

        return times

    if arg is None:
        return arg
    elif isinstance(arg, time):
        return arg
    elif isinstance(arg, ABCSeries):
        values = _convert_listlike(arg._values, format)
        return arg._constructor(values, index=arg.index, name=arg.name)
    elif isinstance(arg, ABCIndex):
        return _convert_listlike(arg, format)
    elif is_list_like(arg):
        return _convert_listlike(arg, format)

    return _convert_listlike(np.array([arg]), format)[0]


# Fixed time formats for time parsing
_time_formats = [
    "%H:%M",
    "%H%M",
    "%I:%M%p",
    "%I%M%p",
    "%H:%M:%S",
    "%H%M%S",
    "%I:%M:%S%p",
    "%I%M%S%p",
]


def _guess_time_format_for_array(arr):
    # Try to guess the format based on the first non-NaN element
    non_nan_elements = notna(arr).nonzero()[0]
    if len(non_nan_elements):
        element = arr[non_nan_elements[0]]
        for time_format in _time_formats:
            try:
                datetime.strptime(element, time_format)
                return time_format
            except ValueError:
                pass

    return None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\cachecontrol\adapter.py ===
# SPDX-FileCopyrightText: 2015 Eric Larson
#
# SPDX-License-Identifier: Apache-2.0
from __future__ import annotations

import functools
import types
import weakref
import zlib
from typing import TYPE_CHECKING, Any, Collection, Mapping

from pip._vendor.requests.adapters import HTTPAdapter

from pip._vendor.cachecontrol.cache import DictCache
from pip._vendor.cachecontrol.controller import PERMANENT_REDIRECT_STATUSES, CacheController
from pip._vendor.cachecontrol.filewrapper import CallbackFileWrapper

if TYPE_CHECKING:
    from pip._vendor.requests import PreparedRequest, Response
    from pip._vendor.urllib3 import HTTPResponse

    from pip._vendor.cachecontrol.cache import BaseCache
    from pip._vendor.cachecontrol.heuristics import BaseHeuristic
    from pip._vendor.cachecontrol.serialize import Serializer


class CacheControlAdapter(HTTPAdapter):
    invalidating_methods = {"PUT", "PATCH", "DELETE"}

    def __init__(
        self,
        cache: BaseCache | None = None,
        cache_etags: bool = True,
        controller_class: type[CacheController] | None = None,
        serializer: Serializer | None = None,
        heuristic: BaseHeuristic | None = None,
        cacheable_methods: Collection[str] | None = None,
        *args: Any,
        **kw: Any,
    ) -> None:
        super().__init__(*args, **kw)
        self.cache = DictCache() if cache is None else cache
        self.heuristic = heuristic
        self.cacheable_methods = cacheable_methods or ("GET",)

        controller_factory = controller_class or CacheController
        self.controller = controller_factory(
            self.cache, cache_etags=cache_etags, serializer=serializer
        )

    def send(
        self,
        request: PreparedRequest,
        stream: bool = False,
        timeout: None | float | tuple[float, float] | tuple[float, None] = None,
        verify: bool | str = True,
        cert: (None | bytes | str | tuple[bytes | str, bytes | str]) = None,
        proxies: Mapping[str, str] | None = None,
        cacheable_methods: Collection[str] | None = None,
    ) -> Response:
        """
        Send a request. Use the request information to see if it
        exists in the cache and cache the response if we need to and can.
        """
        cacheable = cacheable_methods or self.cacheable_methods
        if request.method in cacheable:
            try:
                cached_response = self.controller.cached_request(request)
            except zlib.error:
                cached_response = None
            if cached_response:
                return self.build_response(request, cached_response, from_cache=True)

            # check for etags and add headers if appropriate
            request.headers.update(self.controller.conditional_headers(request))

        resp = super().send(request, stream, timeout, verify, cert, proxies)

        return resp

    def build_response(  # type: ignore[override]
        self,
        request: PreparedRequest,
        response: HTTPResponse,
        from_cache: bool = False,
        cacheable_methods: Collection[str] | None = None,
    ) -> Response:
        """
        Build a response by making a request or using the cache.

        This will end up calling send and returning a potentially
        cached response
        """
        cacheable = cacheable_methods or self.cacheable_methods
        if not from_cache and request.method in cacheable:
            # Check for any heuristics that might update headers
            # before trying to cache.
            if self.heuristic:
                response = self.heuristic.apply(response)

            # apply any expiration heuristics
            if response.status == 304:
                # We must have sent an ETag request. This could mean
                # that we've been expired already or that we simply
                # have an etag. In either case, we want to try and
                # update the cache if that is the case.
                cached_response = self.controller.update_cached_response(
                    request, response
                )

                if cached_response is not response:
                    from_cache = True

                # We are done with the server response, read a
                # possible response body (compliant servers will
                # not return one, but we cannot be 100% sure) and
                # release the connection back to the pool.
                response.read(decode_content=False)
                response.release_conn()

                response = cached_response

            # We always cache the 301 responses
            elif int(response.status) in PERMANENT_REDIRECT_STATUSES:
                self.controller.cache_response(request, response)
            else:
                # Wrap the response file with a wrapper that will cache the
                #   response when the stream has been consumed.
                response._fp = CallbackFileWrapper(  # type: ignore[assignment]
                    response._fp,  # type: ignore[arg-type]
                    functools.partial(
                        self.controller.cache_response, request, weakref.ref(response)
                    ),
                )
                if response.chunked:
                    super_update_chunk_length = response.__class__._update_chunk_length

                    def _update_chunk_length(
                        weak_self: weakref.ReferenceType[HTTPResponse],
                    ) -> None:
                        self = weak_self()
                        if self is None:
                            return

                        super_update_chunk_length(self)
                        if self.chunk_left == 0:
                            self._fp._close()  # type: ignore[union-attr]

                    response._update_chunk_length = functools.partial(  # type: ignore[method-assign]
                        _update_chunk_length, weakref.ref(response)
                    )

        resp: Response = super().build_response(request, response)

        # See if we should invalidate the cache.
        if request.method in self.invalidating_methods and resp.ok:
            assert request.url is not None
            cache_url = self.controller.cache_url(request.url)
            self.cache.delete(cache_url)

        # Give the request a from_cache attr to let people use it
        resp.from_cache = from_cache  # type: ignore[attr-defined]

        return resp

    def close(self) -> None:
        self.cache.close()
        super().close()  # type: ignore[no-untyped-call]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\soupsieve\__init__.py ===
"""
Soup Sieve.

A CSS selector filter for BeautifulSoup4.

MIT License

Copyright (c) 2018 Isaac Muse

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
from __future__ import annotations
from .__meta__ import __version__, __version_info__  # noqa: F401
from . import css_parser as cp
from . import css_match as cm
from . import css_types as ct
from .util import DEBUG, SelectorSyntaxError  # noqa: F401
import bs4
from typing import Any, Iterator, Iterable

__all__ = (
    'DEBUG', 'SelectorSyntaxError', 'SoupSieve',
    'closest', 'compile', 'filter', 'iselect',
    'match', 'select', 'select_one'
)

SoupSieve = cm.SoupSieve


def compile(  # noqa: A001
    pattern: str,
    namespaces: dict[str, str] | None = None,
    flags: int = 0,
    *,
    custom: dict[str, str] | None = None,
    **kwargs: Any
) -> cm.SoupSieve:
    """Compile CSS pattern."""

    if isinstance(pattern, SoupSieve):
        if flags:
            raise ValueError("Cannot process 'flags' argument on a compiled selector list")
        elif namespaces is not None:
            raise ValueError("Cannot process 'namespaces' argument on a compiled selector list")
        elif custom is not None:
            raise ValueError("Cannot process 'custom' argument on a compiled selector list")
        return pattern

    return cp._cached_css_compile(
        pattern,
        ct.Namespaces(namespaces) if namespaces is not None else namespaces,
        ct.CustomSelectors(custom) if custom is not None else custom,
        flags
    )


def purge() -> None:
    """Purge cached patterns."""

    cp._purge_cache()


def closest(
    select: str,
    tag: bs4.Tag,
    namespaces: dict[str, str] | None = None,
    flags: int = 0,
    *,
    custom: dict[str, str] | None = None,
    **kwargs: Any
) -> bs4.Tag | None:
    """Match closest ancestor."""

    return compile(select, namespaces, flags, **kwargs).closest(tag)


def match(
    select: str,
    tag: bs4.Tag,
    namespaces: dict[str, str] | None = None,
    flags: int = 0,
    *,
    custom: dict[str, str] | None = None,
    **kwargs: Any
) -> bool:
    """Match node."""

    return compile(select, namespaces, flags, **kwargs).match(tag)


def filter(  # noqa: A001
    select: str,
    iterable: Iterable[bs4.Tag],
    namespaces: dict[str, str] | None = None,
    flags: int = 0,
    *,
    custom: dict[str, str] | None = None,
    **kwargs: Any
) -> list[bs4.Tag]:
    """Filter list of nodes."""

    return compile(select, namespaces, flags, **kwargs).filter(iterable)


def select_one(
    select: str,
    tag: bs4.Tag,
    namespaces: dict[str, str] | None = None,
    flags: int = 0,
    *,
    custom: dict[str, str] | None = None,
    **kwargs: Any
) -> bs4.Tag | None:
    """Select a single tag."""

    return compile(select, namespaces, flags, **kwargs).select_one(tag)


def select(
    select: str,
    tag: bs4.Tag,
    namespaces: dict[str, str] | None = None,
    limit: int = 0,
    flags: int = 0,
    *,
    custom: dict[str, str] | None = None,
    **kwargs: Any
) -> list[bs4.Tag]:
    """Select the specified tags."""

    return compile(select, namespaces, flags, **kwargs).select(tag, limit)


def iselect(
    select: str,
    tag: bs4.Tag,
    namespaces: dict[str, str] | None = None,
    limit: int = 0,
    flags: int = 0,
    *,
    custom: dict[str, str] | None = None,
    **kwargs: Any
) -> Iterator[bs4.Tag]:
    """Iterate the specified tags."""

    yield from compile(select, namespaces, flags, **kwargs).iselect(tag, limit)


def escape(ident: str) -> str:
    """Escape identifier."""

    return cp.escape(ident)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\plotting\textplot.py ===
from sympy.core.numbers import Float
from sympy.core.symbol import Dummy
from sympy.utilities.lambdify import lambdify

import math


def is_valid(x):
    """Check if a floating point number is valid"""
    if x is None:
        return False
    if isinstance(x, complex):
        return False
    return not math.isinf(x) and not math.isnan(x)


def rescale(y, W, H, mi, ma):
    """Rescale the given array `y` to fit into the integer values
    between `0` and `H-1` for the values between ``mi`` and ``ma``.
    """
    y_new = []

    norm = ma - mi
    offset = (ma + mi) / 2

    for x in range(W):
        if is_valid(y[x]):
            normalized = (y[x] - offset) / norm
            if not is_valid(normalized):
                y_new.append(None)
            else:
                rescaled = Float((normalized*H + H/2) * (H-1)/H).round()
                rescaled = int(rescaled)
                y_new.append(rescaled)
        else:
            y_new.append(None)
    return y_new


def linspace(start, stop, num):
    return [start + (stop - start) * x / (num-1) for x in range(num)]


def textplot_str(expr, a, b, W=55, H=21):
    """Generator for the lines of the plot"""
    free = expr.free_symbols
    if len(free) > 1:
        raise ValueError(
            "The expression must have a single variable. (Got {})"
            .format(free))
    x = free.pop() if free else Dummy()
    f = lambdify([x], expr)
    if isinstance(a, complex):
        if a.imag == 0:
            a = a.real
    if isinstance(b, complex):
        if b.imag == 0:
            b = b.real
    a = float(a)
    b = float(b)

    # Calculate function values
    x = linspace(a, b, W)
    y = []
    for val in x:
        try:
            y.append(f(val))
        # Not sure what exceptions to catch here or why...
        except (ValueError, TypeError, ZeroDivisionError):
            y.append(None)

    # Normalize height to screen space
    y_valid = list(filter(is_valid, y))
    if y_valid:
        ma = max(y_valid)
        mi = min(y_valid)
        if ma == mi:
            if ma:
                mi, ma = sorted([0, 2*ma])
            else:
                mi, ma = -1, 1
    else:
        mi, ma = -1, 1
    y_range = ma - mi
    precision = math.floor(math.log10(y_range)) - 1
    precision *= -1
    mi = round(mi, precision)
    ma = round(ma, precision)
    y = rescale(y, W, H, mi, ma)

    y_bins = linspace(mi, ma, H)

    # Draw plot
    margin = 7
    for h in range(H - 1, -1, -1):
        s = [' '] * W
        for i in range(W):
            if y[i] == h:
                if (i == 0 or y[i - 1] == h - 1) and (i == W - 1 or y[i + 1] == h + 1):
                    s[i] = '/'
                elif (i == 0 or y[i - 1] == h + 1) and (i == W - 1 or y[i + 1] == h - 1):
                    s[i] = '\\'
                else:
                    s[i] = '.'

        if h == 0:
            for i in range(W):
                s[i] = '_'

        # Print y values
        if h in (0, H//2, H - 1):
            prefix = ("%g" % y_bins[h]).rjust(margin)[:margin]
        else:
            prefix = " "*margin
        s = "".join(s)
        if h == H//2:
            s = s.replace(" ", "-")
        yield prefix + " |" + s

    # Print x values
    bottom = " " * (margin + 2)
    bottom += ("%g" % x[0]).ljust(W//2)
    if W % 2 == 1:
        bottom += ("%g" % x[W//2]).ljust(W//2)
    else:
        bottom += ("%g" % x[W//2]).ljust(W//2-1)
    bottom += "%g" % x[-1]
    yield bottom


def textplot(expr, a, b, W=55, H=21):
    r"""
    Print a crude ASCII art plot of the SymPy expression 'expr' (which
    should contain a single symbol, e.g. x or something else) over the
    interval [a, b].

    Examples
    ========

    >>> from sympy import Symbol, sin
    >>> from sympy.plotting import textplot
    >>> t = Symbol('t')
    >>> textplot(sin(t)*t, 0, 15)
     14 |                                                  ...
        |                                                     .
        |                                                 .
        |                                                      .
        |                                                .
        |                            ...
        |                           /   .               .
        |                          /
        |                         /      .
        |                        .        .            .
    1.5 |----.......--------------------------------------------
        |....       \           .          .
        |            \         /                      .
        |             ..      /             .
        |               \    /                       .
        |                ....
        |                                    .
        |                                     .     .
        |
        |                                      .   .
    -11 |_______________________________________________________
         0                          7.5                        15
    """
    for line in textplot_str(expr, a, b, W, H):
        print(line)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_deprecate.py ===
from __future__ import annotations

import sys
import warnings
from functools import wraps
from typing import TYPE_CHECKING, ClassVar, TypeVar

import attrs

if TYPE_CHECKING:
    from collections.abc import Callable

    from typing_extensions import ParamSpec

    ArgsT = ParamSpec("ArgsT")

RetT = TypeVar("RetT")


# We want our warnings to be visible by default (at least for now), but we
# also want it to be possible to override that using the -W switch. AFAICT
# this means we cannot inherit from DeprecationWarning, because the only way
# to make it visible by default then would be to add our own filter at import
# time, but that would override -W switches...
class TrioDeprecationWarning(FutureWarning):
    """Warning emitted if you use deprecated Trio functionality.

    While a relatively mature project, Trio remains committed to refining its
    design and improving usability. As part of this, we occasionally deprecate
    or remove functionality that proves suboptimal. If you use Trio, we
    recommend `subscribing to issue #1
    <https://github.com/python-trio/trio/issues/1>`__ to get information about
    upcoming deprecations and other backwards compatibility breaking changes.

    Despite the name, this class currently inherits from
    :class:`FutureWarning`, not :class:`DeprecationWarning`, because until a
    1.0 release, we want these warnings to be visible by default. You can hide
    them by installing a filter or with the ``-W`` switch: see the
    :mod:`warnings` documentation for details.
    """


def _url_for_issue(issue: int) -> str:
    return f"https://github.com/python-trio/trio/issues/{issue}"


def _stringify(thing: object) -> str:
    if hasattr(thing, "__module__") and hasattr(thing, "__qualname__"):
        return f"{thing.__module__}.{thing.__qualname__}"
    return str(thing)


def warn_deprecated(
    thing: object,
    version: str,
    *,
    issue: int | None,
    instead: object,
    stacklevel: int = 2,
    use_triodeprecationwarning: bool = False,
) -> None:
    stacklevel += 1
    msg = f"{_stringify(thing)} is deprecated since Trio {version}"
    if instead is None:
        msg += " with no replacement"
    else:
        msg += f"; use {_stringify(instead)} instead"
    if issue is not None:
        msg += f" ({_url_for_issue(issue)})"
    if use_triodeprecationwarning:
        warning_class: type[Warning] = TrioDeprecationWarning
    else:
        warning_class = DeprecationWarning
    warnings.warn(warning_class(msg), stacklevel=stacklevel)


# @deprecated("0.2.0", issue=..., instead=...)
# def ...
def deprecated(
    version: str,
    *,
    thing: object = None,
    issue: int | None,
    instead: object,
    use_triodeprecationwarning: bool = False,
) -> Callable[[Callable[ArgsT, RetT]], Callable[ArgsT, RetT]]:
    def do_wrap(fn: Callable[ArgsT, RetT]) -> Callable[ArgsT, RetT]:
        nonlocal thing

        @wraps(fn)
        def wrapper(*args: ArgsT.args, **kwargs: ArgsT.kwargs) -> RetT:
            warn_deprecated(
                thing,
                version,
                instead=instead,
                issue=issue,
                use_triodeprecationwarning=use_triodeprecationwarning,
            )
            return fn(*args, **kwargs)

        # If our __module__ or __qualname__ get modified, we want to pick up
        # on that, so we read them off the wrapper object instead of the (now
        # hidden) fn object
        if thing is None:
            thing = wrapper

        if wrapper.__doc__ is not None:
            doc = wrapper.__doc__
            doc = doc.rstrip()
            doc += "\n\n"
            doc += f".. deprecated:: {version}\n"
            if instead is not None:
                doc += f"   Use {_stringify(instead)} instead.\n"
            if issue is not None:
                doc += f"   For details, see `issue #{issue} <{_url_for_issue(issue)}>`__.\n"
            doc += "\n"
            wrapper.__doc__ = doc

        return wrapper

    return do_wrap


def deprecated_alias(
    old_qualname: str,
    new_fn: Callable[ArgsT, RetT],
    version: str,
    *,
    issue: int | None,
) -> Callable[ArgsT, RetT]:
    @deprecated(version, issue=issue, instead=new_fn)
    @wraps(new_fn, assigned=("__module__", "__annotations__"))
    def wrapper(*args: ArgsT.args, **kwargs: ArgsT.kwargs) -> RetT:
        """Deprecated alias."""
        return new_fn(*args, **kwargs)

    wrapper.__qualname__ = old_qualname
    wrapper.__name__ = old_qualname.rpartition(".")[-1]
    return wrapper


@attrs.frozen(slots=False)
class DeprecatedAttribute:
    _not_set: ClassVar[object] = object()

    value: object
    version: str
    issue: int | None
    instead: object = _not_set


def deprecate_attributes(
    module_name: str, deprecated_attributes: dict[str, DeprecatedAttribute]
) -> None:
    def __getattr__(name: str) -> object:
        if name in deprecated_attributes:
            info = deprecated_attributes[name]
            instead = info.instead
            if instead is DeprecatedAttribute._not_set:
                instead = info.value
            thing = f"{module_name}.{name}"
            warn_deprecated(thing, info.version, issue=info.issue, instead=instead)
            return info.value

        msg = "module '{}' has no attribute '{}'"
        raise AttributeError(msg.format(module_name, name))

    sys.modules[module_name].__getattr__ = __getattr__  # type: ignore[method-assign]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\dateutil\zoneinfo\__init__.py ===
# -*- coding: utf-8 -*-
import warnings
import json

from tarfile import TarFile
from pkgutil import get_data
from io import BytesIO

from dateutil.tz import tzfile as _tzfile

__all__ = ["get_zonefile_instance", "gettz", "gettz_db_metadata"]

ZONEFILENAME = "dateutil-zoneinfo.tar.gz"
METADATA_FN = 'METADATA'


class tzfile(_tzfile):
    def __reduce__(self):
        return (gettz, (self._filename,))


def getzoneinfofile_stream():
    try:
        return BytesIO(get_data(__name__, ZONEFILENAME))
    except IOError as e:  # TODO  switch to FileNotFoundError?
        warnings.warn("I/O error({0}): {1}".format(e.errno, e.strerror))
        return None


class ZoneInfoFile(object):
    def __init__(self, zonefile_stream=None):
        if zonefile_stream is not None:
            with TarFile.open(fileobj=zonefile_stream) as tf:
                self.zones = {zf.name: tzfile(tf.extractfile(zf), filename=zf.name)
                              for zf in tf.getmembers()
                              if zf.isfile() and zf.name != METADATA_FN}
                # deal with links: They'll point to their parent object. Less
                # waste of memory
                links = {zl.name: self.zones[zl.linkname]
                         for zl in tf.getmembers() if
                         zl.islnk() or zl.issym()}
                self.zones.update(links)
                try:
                    metadata_json = tf.extractfile(tf.getmember(METADATA_FN))
                    metadata_str = metadata_json.read().decode('UTF-8')
                    self.metadata = json.loads(metadata_str)
                except KeyError:
                    # no metadata in tar file
                    self.metadata = None
        else:
            self.zones = {}
            self.metadata = None

    def get(self, name, default=None):
        """
        Wrapper for :func:`ZoneInfoFile.zones.get`. This is a convenience method
        for retrieving zones from the zone dictionary.

        :param name:
            The name of the zone to retrieve. (Generally IANA zone names)

        :param default:
            The value to return in the event of a missing key.

        .. versionadded:: 2.6.0

        """
        return self.zones.get(name, default)


# The current API has gettz as a module function, although in fact it taps into
# a stateful class. So as a workaround for now, without changing the API, we
# will create a new "global" class instance the first time a user requests a
# timezone. Ugly, but adheres to the api.
#
# TODO: Remove after deprecation period.
_CLASS_ZONE_INSTANCE = []


def get_zonefile_instance(new_instance=False):
    """
    This is a convenience function which provides a :class:`ZoneInfoFile`
    instance using the data provided by the ``dateutil`` package. By default, it
    caches a single instance of the ZoneInfoFile object and returns that.

    :param new_instance:
        If ``True``, a new instance of :class:`ZoneInfoFile` is instantiated and
        used as the cached instance for the next call. Otherwise, new instances
        are created only as necessary.

    :return:
        Returns a :class:`ZoneInfoFile` object.

    .. versionadded:: 2.6
    """
    if new_instance:
        zif = None
    else:
        zif = getattr(get_zonefile_instance, '_cached_instance', None)

    if zif is None:
        zif = ZoneInfoFile(getzoneinfofile_stream())

        get_zonefile_instance._cached_instance = zif

    return zif


def gettz(name):
    """
    This retrieves a time zone from the local zoneinfo tarball that is packaged
    with dateutil.

    :param name:
        An IANA-style time zone name, as found in the zoneinfo file.

    :return:
        Returns a :class:`dateutil.tz.tzfile` time zone object.

    .. warning::
        It is generally inadvisable to use this function, and it is only
        provided for API compatibility with earlier versions. This is *not*
        equivalent to ``dateutil.tz.gettz()``, which selects an appropriate
        time zone based on the inputs, favoring system zoneinfo. This is ONLY
        for accessing the dateutil-specific zoneinfo (which may be out of
        date compared to the system zoneinfo).

    .. deprecated:: 2.6
        If you need to use a specific zoneinfofile over the system zoneinfo,
        instantiate a :class:`dateutil.zoneinfo.ZoneInfoFile` object and call
        :func:`dateutil.zoneinfo.ZoneInfoFile.get(name)` instead.

        Use :func:`get_zonefile_instance` to retrieve an instance of the
        dateutil-provided zoneinfo.
    """
    warnings.warn("zoneinfo.gettz() will be removed in future versions, "
                  "to use the dateutil-provided zoneinfo files, instantiate a "
                  "ZoneInfoFile object and use ZoneInfoFile.zones.get() "
                  "instead. See the documentation for details.",
                  DeprecationWarning)

    if len(_CLASS_ZONE_INSTANCE) == 0:
        _CLASS_ZONE_INSTANCE.append(ZoneInfoFile(getzoneinfofile_stream()))
    return _CLASS_ZONE_INSTANCE[0].zones.get(name)


def gettz_db_metadata():
    """ Get the zonefile metadata

    See `zonefile_metadata`_

    :returns:
        A dictionary with the database metadata

    .. deprecated:: 2.6
        See deprecation warning in :func:`zoneinfo.gettz`. To get metadata,
        query the attribute ``zoneinfo.ZoneInfoFile.metadata``.
    """
    warnings.warn("zoneinfo.gettz_db_metadata() will be removed in future "
                  "versions, to use the dateutil-provided zoneinfo files, "
                  "ZoneInfoFile object and query the 'metadata' attribute "
                  "instead. See the documentation for details.",
                  DeprecationWarning)

    if len(_CLASS_ZONE_INSTANCE) == 0:
        _CLASS_ZONE_INSTANCE.append(ZoneInfoFile(getzoneinfofile_stream()))
    return _CLASS_ZONE_INSTANCE[0].metadata

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_transpose.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

from logging import getLogger

from fusion_base import Fusion
from fusion_utils import FusionUtils
from onnx import NodeProto, TensorProto, helper
from onnx_model import OnnxModel

logger = getLogger(__name__)


class FusionTranspose(Fusion):
    def __init__(self, model: OnnxModel):
        super().__init__(model, "Transpose", "Transpose")

    def fuse(
        self,
        transpose_node: NodeProto,
        input_name_to_nodes: dict[str, list[NodeProto]],
        output_name_to_node: dict[str, NodeProto],
    ):
        """
        Note that onnxruntime will do comprehensive transpose optimization after loading model.
        The purpose of this fusion is to make graph clean before running onnxruntime.

        Case 1:
              (input)-->Transpose(perm=a)-->Transpose(perm=b)-->
        After:
              (input)-->Transpose(perm=a)-->  (this path can be removed if the output is not used anymore)
                |
                +----->Transpose(perm=a*b)-->

        Case 2 (Cast has only one child):
              (input)-->Transpose(perm=a)--> Cast -->Transpose(perm=b)-->
        After:
              (input)-->Transpose(perm=a)-->  (this path can be removed if the output is not used anymore)
                |
                +----->Cast --> Transpose(perm=a*b)-->
        """
        transpose_b = transpose_node
        if transpose_b.input[0] not in output_name_to_node:
            return

        transpose_a = output_name_to_node[transpose_b.input[0]]
        if transpose_a.op_type != "Cast":
            cast_node = None
        else:
            cast_node = transpose_a

            cast_children = self.model.get_children(cast_node, input_name_to_nodes)
            if cast_children and len(cast_children) > 1:
                return

            if cast_node.input[0] not in output_name_to_node:
                return

            transpose_a = output_name_to_node[cast_node.input[0]]

        if transpose_a.op_type != "Transpose":
            return

        permutation = OnnxModel.get_node_attribute(transpose_b, "perm")
        assert isinstance(permutation, list)

        parent_permutation = OnnxModel.get_node_attribute(transpose_a, "perm")
        assert isinstance(parent_permutation, list)

        assert len(parent_permutation) == len(permutation)

        output_permutation = []
        for _j, index in enumerate(permutation):
            output_permutation.append(parent_permutation[index])

        if cast_node is None:
            if FusionUtils.skip_parent(self.model, transpose_b, transpose_a, input_name_to_nodes):
                self.nodes_to_remove.append(transpose_a)
        else:
            if FusionUtils.skip_parent(self.model, cast_node, transpose_a, input_name_to_nodes):
                self.nodes_to_remove.append(transpose_a)
        transpose_b.ClearField("attribute")
        transpose_b.attribute.extend([helper.make_attribute("perm", output_permutation)])


class FusionInsertTranspose(Fusion):
    def __init__(self, model: OnnxModel):
        super().__init__(model, "", "GroupNorm")

    def create_transpose_node(self, input_name: str, perm: list[int], output_name=None):
        """Append a Transpose node after an input"""
        node_name = self.model.create_node_name("Transpose")
        if output_name is None:
            output_name = node_name + "_out" + "-" + input_name
        transpose_node = helper.make_node("Transpose", inputs=[input_name], outputs=[output_name], name=node_name)
        transpose_node.attribute.extend([helper.make_attribute("perm", perm)])
        return transpose_node

    def fuse(
        self,
        group_norm_node: NodeProto,
        input_name_to_nodes: dict[str, list[NodeProto]],
        output_name_to_node: dict[str, NodeProto],
    ):
        """
        This optimization will insert an Transpose, and onnxruntime transpose optimizer will remove it together with
        another Transpose so that we can get effect of reducing one Transpose after onnxruntime optimization.
        Before:
            --> Gemm --> Unsqueeze(axes=[2]) --> Unsqueeze(axes=[3]) --> Add --> Transpose([0,2,3,1]) --> GroupNorm
        After:
            --> Gemm --> Unsqueeze(axes=[1]) --> Unsqueeze(axes=[2]) -->Transpose([0,3,1,2]) --> Add --> Transpose([0,2,3,1]) --> GroupNorm
        """
        gemm_path = self.model.match_parent_path(
            group_norm_node, ["Transpose", "Add", "Unsqueeze", "Unsqueeze", "Gemm"], [0, 0, None, 0, 0]
        )
        if gemm_path is None:
            return
        transpose, add, unsqueeze_3, unsqueeze_2, gemm = gemm_path
        if self.model.find_graph_output(unsqueeze_3.output[0]):
            return

        permutation = OnnxModel.get_node_attribute(transpose, "perm")
        assert isinstance(permutation, list)
        if permutation != [0, 2, 3, 1]:
            return

        if not (
            len(unsqueeze_3.input) == 2
            and self.model.get_constant_value(unsqueeze_3.input[1]) == 3
            and len(unsqueeze_2.input) == 2
            and self.model.get_constant_value(unsqueeze_2.input[1]) == 2
            and len(self.model.get_children(gemm, input_name_to_nodes)) == 1
            and len(self.model.get_children(unsqueeze_3, input_name_to_nodes)) == 1
            and len(self.model.get_children(unsqueeze_2, input_name_to_nodes)) == 1
        ):
            return

        # Here we use hard-coded name so that it could be shared for the whole model.
        axes_1 = "ort_const_unsqueeze_axes_1"
        if self.model.get_initializer(axes_1) is None:
            self.add_initializer(
                name=axes_1,
                data_type=TensorProto.INT64,
                dims=[1],
                vals=[1],
                raw=False,
            )

        axes_2 = "ort_const_unsqueeze_axes_2"
        if self.model.get_initializer(axes_2) is None:
            self.add_initializer(
                name=axes_2,
                data_type=TensorProto.INT64,
                dims=[1],
                vals=[2],
                raw=False,
            )

        unsqueeze_3.input[1] = "ort_const_unsqueeze_axes_2"
        unsqueeze_2.input[1] = "ort_const_unsqueeze_axes_1"
        transpose_output_name = self.model.create_node_name("Transpose") + "_NCHW"
        self.model.replace_input_of_all_nodes(unsqueeze_3.output[0], transpose_output_name)
        new_transpose = self.create_transpose_node(unsqueeze_3.output[0], [0, 3, 1, 2], transpose_output_name)
        self.model.add_node(new_transpose, self.this_graph_name)
        self.increase_counter("Insert Transpose")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\containers.py ===
from itertools import zip_longest
from typing import (
    TYPE_CHECKING,
    Iterable,
    Iterator,
    List,
    Optional,
    TypeVar,
    Union,
    overload,
)

if TYPE_CHECKING:
    from .console import (
        Console,
        ConsoleOptions,
        JustifyMethod,
        OverflowMethod,
        RenderResult,
        RenderableType,
    )
    from .text import Text

from .cells import cell_len
from .measure import Measurement

T = TypeVar("T")


class Renderables:
    """A list subclass which renders its contents to the console."""

    def __init__(
        self, renderables: Optional[Iterable["RenderableType"]] = None
    ) -> None:
        self._renderables: List["RenderableType"] = (
            list(renderables) if renderables is not None else []
        )

    def __rich_console__(
        self, console: "Console", options: "ConsoleOptions"
    ) -> "RenderResult":
        """Console render method to insert line-breaks."""
        yield from self._renderables

    def __rich_measure__(
        self, console: "Console", options: "ConsoleOptions"
    ) -> "Measurement":
        dimensions = [
            Measurement.get(console, options, renderable)
            for renderable in self._renderables
        ]
        if not dimensions:
            return Measurement(1, 1)
        _min = max(dimension.minimum for dimension in dimensions)
        _max = max(dimension.maximum for dimension in dimensions)
        return Measurement(_min, _max)

    def append(self, renderable: "RenderableType") -> None:
        self._renderables.append(renderable)

    def __iter__(self) -> Iterable["RenderableType"]:
        return iter(self._renderables)


class Lines:
    """A list subclass which can render to the console."""

    def __init__(self, lines: Iterable["Text"] = ()) -> None:
        self._lines: List["Text"] = list(lines)

    def __repr__(self) -> str:
        return f"Lines({self._lines!r})"

    def __iter__(self) -> Iterator["Text"]:
        return iter(self._lines)

    @overload
    def __getitem__(self, index: int) -> "Text":
        ...

    @overload
    def __getitem__(self, index: slice) -> List["Text"]:
        ...

    def __getitem__(self, index: Union[slice, int]) -> Union["Text", List["Text"]]:
        return self._lines[index]

    def __setitem__(self, index: int, value: "Text") -> "Lines":
        self._lines[index] = value
        return self

    def __len__(self) -> int:
        return self._lines.__len__()

    def __rich_console__(
        self, console: "Console", options: "ConsoleOptions"
    ) -> "RenderResult":
        """Console render method to insert line-breaks."""
        yield from self._lines

    def append(self, line: "Text") -> None:
        self._lines.append(line)

    def extend(self, lines: Iterable["Text"]) -> None:
        self._lines.extend(lines)

    def pop(self, index: int = -1) -> "Text":
        return self._lines.pop(index)

    def justify(
        self,
        console: "Console",
        width: int,
        justify: "JustifyMethod" = "left",
        overflow: "OverflowMethod" = "fold",
    ) -> None:
        """Justify and overflow text to a given width.

        Args:
            console (Console): Console instance.
            width (int): Number of cells available per line.
            justify (str, optional): Default justify method for text: "left", "center", "full" or "right". Defaults to "left".
            overflow (str, optional): Default overflow for text: "crop", "fold", or "ellipsis". Defaults to "fold".

        """
        from .text import Text

        if justify == "left":
            for line in self._lines:
                line.truncate(width, overflow=overflow, pad=True)
        elif justify == "center":
            for line in self._lines:
                line.rstrip()
                line.truncate(width, overflow=overflow)
                line.pad_left((width - cell_len(line.plain)) // 2)
                line.pad_right(width - cell_len(line.plain))
        elif justify == "right":
            for line in self._lines:
                line.rstrip()
                line.truncate(width, overflow=overflow)
                line.pad_left(width - cell_len(line.plain))
        elif justify == "full":
            for line_index, line in enumerate(self._lines):
                if line_index == len(self._lines) - 1:
                    break
                words = line.split(" ")
                words_size = sum(cell_len(word.plain) for word in words)
                num_spaces = len(words) - 1
                spaces = [1 for _ in range(num_spaces)]
                index = 0
                if spaces:
                    while words_size + num_spaces < width:
                        spaces[len(spaces) - index - 1] += 1
                        num_spaces += 1
                        index = (index + 1) % len(spaces)
                tokens: List[Text] = []
                for index, (word, next_word) in enumerate(
                    zip_longest(words, words[1:])
                ):
                    tokens.append(word)
                    if index < len(spaces):
                        style = word.get_style_at_offset(console, -1)
                        next_style = next_word.get_style_at_offset(console, 0)
                        space_style = style if style == next_style else line.style
                        tokens.append(Text(" " * spaces[index], style=space_style))
                self[line_index] = Text("").join(tokens)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\postgresql\__init__.py ===
# dialects/postgresql/__init__.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors

from types import ModuleType

from . import array as arraylib  # noqa # keep above base and other dialects
from . import asyncpg  # noqa
from . import base
from . import pg8000  # noqa
from . import psycopg  # noqa
from . import psycopg2  # noqa
from . import psycopg2cffi  # noqa
from .array import All
from .array import Any
from .array import ARRAY
from .array import array
from .base import BIGINT
from .base import BOOLEAN
from .base import CHAR
from .base import DATE
from .base import DOMAIN
from .base import DOUBLE_PRECISION
from .base import FLOAT
from .base import INTEGER
from .base import NUMERIC
from .base import REAL
from .base import SMALLINT
from .base import TEXT
from .base import UUID
from .base import VARCHAR
from .dml import Insert
from .dml import insert
from .ext import aggregate_order_by
from .ext import array_agg
from .ext import ExcludeConstraint
from .ext import phraseto_tsquery
from .ext import plainto_tsquery
from .ext import to_tsquery
from .ext import to_tsvector
from .ext import ts_headline
from .ext import websearch_to_tsquery
from .hstore import HSTORE
from .hstore import hstore
from .json import JSON
from .json import JSONB
from .json import JSONPATH
from .named_types import CreateDomainType
from .named_types import CreateEnumType
from .named_types import DropDomainType
from .named_types import DropEnumType
from .named_types import ENUM
from .named_types import NamedType
from .ranges import AbstractMultiRange
from .ranges import AbstractRange
from .ranges import AbstractSingleRange
from .ranges import DATEMULTIRANGE
from .ranges import DATERANGE
from .ranges import INT4MULTIRANGE
from .ranges import INT4RANGE
from .ranges import INT8MULTIRANGE
from .ranges import INT8RANGE
from .ranges import MultiRange
from .ranges import NUMMULTIRANGE
from .ranges import NUMRANGE
from .ranges import Range
from .ranges import TSMULTIRANGE
from .ranges import TSRANGE
from .ranges import TSTZMULTIRANGE
from .ranges import TSTZRANGE
from .types import BIT
from .types import BYTEA
from .types import CIDR
from .types import CITEXT
from .types import INET
from .types import INTERVAL
from .types import MACADDR
from .types import MACADDR8
from .types import MONEY
from .types import OID
from .types import REGCLASS
from .types import REGCONFIG
from .types import TIME
from .types import TIMESTAMP
from .types import TSQUERY
from .types import TSVECTOR


# Alias psycopg also as psycopg_async
psycopg_async = type(
    "psycopg_async", (ModuleType,), {"dialect": psycopg.dialect_async}
)

base.dialect = dialect = psycopg2.dialect


__all__ = (
    "INTEGER",
    "BIGINT",
    "SMALLINT",
    "VARCHAR",
    "CHAR",
    "TEXT",
    "NUMERIC",
    "FLOAT",
    "REAL",
    "INET",
    "CIDR",
    "CITEXT",
    "UUID",
    "BIT",
    "MACADDR",
    "MACADDR8",
    "MONEY",
    "OID",
    "REGCLASS",
    "REGCONFIG",
    "TSQUERY",
    "TSVECTOR",
    "DOUBLE_PRECISION",
    "TIMESTAMP",
    "TIME",
    "DATE",
    "BYTEA",
    "BOOLEAN",
    "INTERVAL",
    "ARRAY",
    "ENUM",
    "DOMAIN",
    "dialect",
    "array",
    "HSTORE",
    "hstore",
    "INT4RANGE",
    "INT8RANGE",
    "NUMRANGE",
    "DATERANGE",
    "INT4MULTIRANGE",
    "INT8MULTIRANGE",
    "NUMMULTIRANGE",
    "DATEMULTIRANGE",
    "TSVECTOR",
    "TSRANGE",
    "TSTZRANGE",
    "TSMULTIRANGE",
    "TSTZMULTIRANGE",
    "JSON",
    "JSONB",
    "JSONPATH",
    "Any",
    "All",
    "DropEnumType",
    "DropDomainType",
    "CreateDomainType",
    "NamedType",
    "CreateEnumType",
    "ExcludeConstraint",
    "Range",
    "aggregate_order_by",
    "array_agg",
    "insert",
    "Insert",
)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\engine\util.py ===
# engine/util.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

from __future__ import annotations

import typing
from typing import Any
from typing import Callable
from typing import Optional
from typing import TypeVar

from .. import exc
from .. import util
from ..util._has_cy import HAS_CYEXTENSION
from ..util.typing import Protocol
from ..util.typing import Self

if typing.TYPE_CHECKING or not HAS_CYEXTENSION:
    from ._py_util import _distill_params_20 as _distill_params_20
    from ._py_util import _distill_raw_params as _distill_raw_params
else:
    from sqlalchemy.cyextension.util import (  # noqa: F401
        _distill_params_20 as _distill_params_20,
    )
    from sqlalchemy.cyextension.util import (  # noqa: F401
        _distill_raw_params as _distill_raw_params,
    )

_C = TypeVar("_C", bound=Callable[[], Any])


def connection_memoize(key: str) -> Callable[[_C], _C]:
    """Decorator, memoize a function in a connection.info stash.

    Only applicable to functions which take no arguments other than a
    connection.  The memo will be stored in ``connection.info[key]``.
    """

    @util.decorator
    def decorated(fn, self, connection):  # type: ignore
        connection = connection.connect()
        try:
            return connection.info[key]
        except KeyError:
            connection.info[key] = val = fn(self, connection)
            return val

    return decorated


class _TConsSubject(Protocol):
    _trans_context_manager: Optional[TransactionalContext]


class TransactionalContext:
    """Apply Python context manager behavior to transaction objects.

    Performs validation to ensure the subject of the transaction is not
    used if the transaction were ended prematurely.

    """

    __slots__ = ("_outer_trans_ctx", "_trans_subject", "__weakref__")

    _trans_subject: Optional[_TConsSubject]

    def _transaction_is_active(self) -> bool:
        raise NotImplementedError()

    def _transaction_is_closed(self) -> bool:
        raise NotImplementedError()

    def _rollback_can_be_called(self) -> bool:
        """indicates the object is in a state that is known to be acceptable
        for rollback() to be called.

        This does not necessarily mean rollback() will succeed or not raise
        an error, just that there is currently no state detected that indicates
        rollback() would fail or emit warnings.

        It also does not mean that there's a transaction in progress, as
        it is usually safe to call rollback() even if no transaction is
        present.

        .. versionadded:: 1.4.28

        """
        raise NotImplementedError()

    def _get_subject(self) -> _TConsSubject:
        raise NotImplementedError()

    def commit(self) -> None:
        raise NotImplementedError()

    def rollback(self) -> None:
        raise NotImplementedError()

    def close(self) -> None:
        raise NotImplementedError()

    @classmethod
    def _trans_ctx_check(cls, subject: _TConsSubject) -> None:
        trans_context = subject._trans_context_manager
        if trans_context:
            if not trans_context._transaction_is_active():
                raise exc.InvalidRequestError(
                    "Can't operate on closed transaction inside context "
                    "manager.  Please complete the context manager "
                    "before emitting further commands."
                )

    def __enter__(self) -> Self:
        subject = self._get_subject()

        # none for outer transaction, may be non-None for nested
        # savepoint, legacy nesting cases
        trans_context = subject._trans_context_manager
        self._outer_trans_ctx = trans_context

        self._trans_subject = subject
        subject._trans_context_manager = self
        return self

    def __exit__(self, type_: Any, value: Any, traceback: Any) -> None:
        subject = getattr(self, "_trans_subject", None)

        # simplistically we could assume that
        # "subject._trans_context_manager is self".  However, any calling
        # code that is manipulating __exit__ directly would break this
        # assumption.  alembic context manager
        # is an example of partial use that just calls __exit__ and
        # not __enter__ at the moment.  it's safe to assume this is being done
        # in the wild also
        out_of_band_exit = (
            subject is None or subject._trans_context_manager is not self
        )

        if type_ is None and self._transaction_is_active():
            try:
                self.commit()
            except:
                with util.safe_reraise():
                    if self._rollback_can_be_called():
                        self.rollback()
            finally:
                if not out_of_band_exit:
                    assert subject is not None
                    subject._trans_context_manager = self._outer_trans_ctx
                self._trans_subject = self._outer_trans_ctx = None
        else:
            try:
                if not self._transaction_is_active():
                    if not self._transaction_is_closed():
                        self.close()
                else:
                    if self._rollback_can_be_called():
                        self.rollback()
            finally:
                if not out_of_band_exit:
                    assert subject is not None
                    subject._trans_context_manager = self._outer_trans_ctx
                self._trans_subject = self._outer_trans_ctx = None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\expressions\trace.py ===
from sympy.core.basic import Basic
from sympy.core.expr import Expr, ExprBuilder
from sympy.core.singleton import S
from sympy.core.sorting import default_sort_key
from sympy.core.symbol import uniquely_named_symbol
from sympy.core.sympify import sympify
from sympy.matrices.matrixbase import MatrixBase
from sympy.matrices.exceptions import NonSquareMatrixError


class Trace(Expr):
    """Matrix Trace

    Represents the trace of a matrix expression.

    Examples
    ========

    >>> from sympy import MatrixSymbol, Trace, eye
    >>> A = MatrixSymbol('A', 3, 3)
    >>> Trace(A)
    Trace(A)
    >>> Trace(eye(3))
    Trace(Matrix([
    [1, 0, 0],
    [0, 1, 0],
    [0, 0, 1]]))
    >>> Trace(eye(3)).simplify()
    3
    """
    is_Trace = True
    is_commutative = True

    def __new__(cls, mat):
        mat = sympify(mat)

        if not mat.is_Matrix:
            raise TypeError("input to Trace, %s, is not a matrix" % str(mat))

        if mat.is_square is False:
            raise NonSquareMatrixError("Trace of a non-square matrix")

        return Basic.__new__(cls, mat)

    def _eval_transpose(self):
        return self

    def _eval_derivative(self, v):
        from sympy.concrete.summations import Sum
        from .matexpr import MatrixElement
        if isinstance(v, MatrixElement):
            return self.rewrite(Sum).diff(v)
        expr = self.doit()
        if isinstance(expr, Trace):
            # Avoid looping infinitely:
            raise NotImplementedError
        return expr._eval_derivative(v)

    def _eval_derivative_matrix_lines(self, x):
        from sympy.tensor.array.expressions.array_expressions import ArrayTensorProduct, ArrayContraction
        r = self.args[0]._eval_derivative_matrix_lines(x)
        for lr in r:
            if lr.higher == 1:
                lr.higher = ExprBuilder(
                    ArrayContraction,
                    [
                        ExprBuilder(
                            ArrayTensorProduct,
                            [
                                lr._lines[0],
                                lr._lines[1],
                            ]
                        ),
                        (1, 3),
                    ],
                    validator=ArrayContraction._validate
                )
            else:
                # This is not a matrix line:
                lr.higher = ExprBuilder(
                    ArrayContraction,
                    [
                        ExprBuilder(
                            ArrayTensorProduct,
                            [
                                lr._lines[0],
                                lr._lines[1],
                                lr.higher,
                            ]
                        ),
                        (1, 3), (0, 2)
                    ]
                )
            lr._lines = [S.One, S.One]
            lr._first_pointer_parent = lr._lines
            lr._second_pointer_parent = lr._lines
            lr._first_pointer_index = 0
            lr._second_pointer_index = 1
        return r

    @property
    def arg(self):
        return self.args[0]

    def doit(self, **hints):
        if hints.get('deep', True):
            arg = self.arg.doit(**hints)
            result = arg._eval_trace()
            if result is not None:
                return result
            else:
                return Trace(arg)
        else:
            # _eval_trace would go too deep here
            if isinstance(self.arg, MatrixBase):
                return trace(self.arg)
            else:
                return Trace(self.arg)

    def as_explicit(self):
        return Trace(self.arg.as_explicit()).doit()

    def _normalize(self):
        # Normalization of trace of matrix products. Use transposition and
        # cyclic properties of traces to make sure the arguments of the matrix
        # product are sorted and the first argument is not a transposition.
        from sympy.matrices.expressions.matmul import MatMul
        from sympy.matrices.expressions.transpose import Transpose
        trace_arg = self.arg
        if isinstance(trace_arg, MatMul):

            def get_arg_key(x):
                a = trace_arg.args[x]
                if isinstance(a, Transpose):
                    a = a.arg
                return default_sort_key(a)

            indmin = min(range(len(trace_arg.args)), key=get_arg_key)
            if isinstance(trace_arg.args[indmin], Transpose):
                trace_arg = Transpose(trace_arg).doit()
                indmin = min(range(len(trace_arg.args)), key=lambda x: default_sort_key(trace_arg.args[x]))
            trace_arg = MatMul.fromiter(trace_arg.args[indmin:] + trace_arg.args[:indmin])
            return Trace(trace_arg)
        return self

    def _eval_rewrite_as_Sum(self, expr, **kwargs):
        from sympy.concrete.summations import Sum
        i = uniquely_named_symbol('i', [expr])
        s = Sum(self.arg[i, i], (i, 0, self.arg.rows - 1))
        return s.doit()


def trace(expr):
    """Trace of a Matrix.  Sum of the diagonal elements.

    Examples
    ========

    >>> from sympy import trace, Symbol, MatrixSymbol, eye
    >>> n = Symbol('n')
    >>> X = MatrixSymbol('X', n, n)  # A square matrix
    >>> trace(2*X)
    2*Trace(X)
    >>> trace(eye(3))
    3
    """
    return Trace(expr).doit()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\stats\sampling\sample_scipy.py ===
from functools import singledispatch

from sympy.core.symbol import Dummy
from sympy.functions.elementary.exponential import exp
from sympy.utilities.lambdify import lambdify
from sympy.external import import_module
from sympy.stats import DiscreteDistributionHandmade
from sympy.stats.crv import SingleContinuousDistribution
from sympy.stats.crv_types import ChiSquaredDistribution, ExponentialDistribution, GammaDistribution, \
    LogNormalDistribution, NormalDistribution, ParetoDistribution, UniformDistribution, BetaDistribution, \
    StudentTDistribution, CauchyDistribution
from sympy.stats.drv_types import GeometricDistribution, LogarithmicDistribution, NegativeBinomialDistribution, \
    PoissonDistribution, SkellamDistribution, YuleSimonDistribution, ZetaDistribution
from sympy.stats.frv import SingleFiniteDistribution


scipy = import_module("scipy", import_kwargs={'fromlist':['stats']})


@singledispatch
def do_sample_scipy(dist, size, seed):
    return None


# CRV

@do_sample_scipy.register(SingleContinuousDistribution)
def _(dist: SingleContinuousDistribution, size, seed):
    # if we don't need to make a handmade pdf, we won't
    import scipy.stats

    z = Dummy('z')
    handmade_pdf = lambdify(z, dist.pdf(z), ['numpy', 'scipy'])

    class scipy_pdf(scipy.stats.rv_continuous):
        def _pdf(dist, x):
            return handmade_pdf(x)

    scipy_rv = scipy_pdf(a=float(dist.set._inf),
                         b=float(dist.set._sup), name='scipy_pdf')
    return scipy_rv.rvs(size=size, random_state=seed)


@do_sample_scipy.register(ChiSquaredDistribution)
def _(dist: ChiSquaredDistribution, size, seed):
    # same parametrisation
    return scipy.stats.chi2.rvs(df=float(dist.k), size=size, random_state=seed)


@do_sample_scipy.register(ExponentialDistribution)
def _(dist: ExponentialDistribution, size, seed):
    # https://docs.scipy.org/doc/scipy/reference/generated/scipy.stats.expon.html#scipy.stats.expon
    return scipy.stats.expon.rvs(scale=1 / float(dist.rate), size=size, random_state=seed)


@do_sample_scipy.register(GammaDistribution)
def _(dist: GammaDistribution, size, seed):
    # https://stackoverflow.com/questions/42150965/how-to-plot-gamma-distribution-with-alpha-and-beta-parameters-in-python
    return scipy.stats.gamma.rvs(a=float(dist.k), scale=float(dist.theta), size=size, random_state=seed)


@do_sample_scipy.register(LogNormalDistribution)
def _(dist: LogNormalDistribution, size, seed):
    # https://docs.scipy.org/doc/scipy/reference/generated/scipy.stats.lognorm.html
    return scipy.stats.lognorm.rvs(scale=float(exp(dist.mean)), s=float(dist.std), size=size, random_state=seed)


@do_sample_scipy.register(NormalDistribution)
def _(dist: NormalDistribution, size, seed):
    return scipy.stats.norm.rvs(loc=float(dist.mean), scale=float(dist.std), size=size, random_state=seed)


@do_sample_scipy.register(ParetoDistribution)
def _(dist: ParetoDistribution, size, seed):
    # https://stackoverflow.com/questions/42260519/defining-pareto-distribution-in-python-scipy
    return scipy.stats.pareto.rvs(b=float(dist.alpha), scale=float(dist.xm), size=size, random_state=seed)


@do_sample_scipy.register(StudentTDistribution)
def _(dist: StudentTDistribution, size, seed):
    return scipy.stats.t.rvs(df=float(dist.nu), size=size, random_state=seed)


@do_sample_scipy.register(UniformDistribution)
def _(dist: UniformDistribution, size, seed):
    # https://docs.scipy.org/doc/scipy/reference/generated/scipy.stats.uniform.html
    return scipy.stats.uniform.rvs(loc=float(dist.left), scale=float(dist.right - dist.left), size=size, random_state=seed)


@do_sample_scipy.register(BetaDistribution)
def _(dist: BetaDistribution, size, seed):
    # same parametrisation
    return scipy.stats.beta.rvs(a=float(dist.alpha), b=float(dist.beta), size=size, random_state=seed)


@do_sample_scipy.register(CauchyDistribution)
def _(dist: CauchyDistribution, size, seed):
    return scipy.stats.cauchy.rvs(loc=float(dist.x0), scale=float(dist.gamma), size=size, random_state=seed)


# DRV:

@do_sample_scipy.register(DiscreteDistributionHandmade)
def _(dist: DiscreteDistributionHandmade, size, seed):
    from scipy.stats import rv_discrete

    z = Dummy('z')
    handmade_pmf = lambdify(z, dist.pdf(z), ['numpy', 'scipy'])

    class scipy_pmf(rv_discrete):
        def _pmf(dist, x):
            return handmade_pmf(x)

    scipy_rv = scipy_pmf(a=float(dist.set._inf), b=float(dist.set._sup),
                         name='scipy_pmf')
    return scipy_rv.rvs(size=size, random_state=seed)


@do_sample_scipy.register(GeometricDistribution)
def _(dist: GeometricDistribution, size, seed):
    return scipy.stats.geom.rvs(p=float(dist.p), size=size, random_state=seed)


@do_sample_scipy.register(LogarithmicDistribution)
def _(dist: LogarithmicDistribution, size, seed):
    return scipy.stats.logser.rvs(p=float(dist.p), size=size, random_state=seed)


@do_sample_scipy.register(NegativeBinomialDistribution)
def _(dist: NegativeBinomialDistribution, size, seed):
    return scipy.stats.nbinom.rvs(n=float(dist.r), p=float(dist.p), size=size, random_state=seed)


@do_sample_scipy.register(PoissonDistribution)
def _(dist: PoissonDistribution, size, seed):
    return scipy.stats.poisson.rvs(mu=float(dist.lamda), size=size, random_state=seed)


@do_sample_scipy.register(SkellamDistribution)
def _(dist: SkellamDistribution, size, seed):
    return scipy.stats.skellam.rvs(mu1=float(dist.mu1), mu2=float(dist.mu2), size=size, random_state=seed)


@do_sample_scipy.register(YuleSimonDistribution)
def _(dist: YuleSimonDistribution, size, seed):
    return scipy.stats.yulesimon.rvs(alpha=float(dist.rho), size=size, random_state=seed)


@do_sample_scipy.register(ZetaDistribution)
def _(dist: ZetaDistribution, size, seed):
    return scipy.stats.zipf.rvs(a=float(dist.s), size=size, random_state=seed)


# FRV:

@do_sample_scipy.register(SingleFiniteDistribution)
def _(dist: SingleFiniteDistribution, size, seed):
    # scipy can handle with custom distributions

    from scipy.stats import rv_discrete
    density_ = dist.dict
    x, y = [], []
    for k, v in density_.items():
        x.append(int(k))
        y.append(float(v))
    scipy_rv = rv_discrete(name='scipy_rv', values=(x, y))
    return scipy_rv.rvs(size=size, random_state=seed)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\werkzeug\sansio\utils.py ===
from __future__ import annotations

import typing as t
from urllib.parse import quote

from .._internal import _plain_int
from ..exceptions import SecurityError
from ..urls import uri_to_iri


def host_is_trusted(hostname: str | None, trusted_list: t.Iterable[str]) -> bool:
    """Check if a host matches a list of trusted names.

    :param hostname: The name to check.
    :param trusted_list: A list of valid names to match. If a name
        starts with a dot it will match all subdomains.

    .. versionadded:: 0.9
    """
    if not hostname:
        return False

    try:
        hostname = hostname.partition(":")[0].encode("idna").decode("ascii")
    except UnicodeEncodeError:
        return False

    if isinstance(trusted_list, str):
        trusted_list = [trusted_list]

    for ref in trusted_list:
        if ref.startswith("."):
            ref = ref[1:]
            suffix_match = True
        else:
            suffix_match = False

        try:
            ref = ref.partition(":")[0].encode("idna").decode("ascii")
        except UnicodeEncodeError:
            return False

        if ref == hostname or (suffix_match and hostname.endswith(f".{ref}")):
            return True

    return False


def get_host(
    scheme: str,
    host_header: str | None,
    server: tuple[str, int | None] | None = None,
    trusted_hosts: t.Iterable[str] | None = None,
) -> str:
    """Return the host for the given parameters.

    This first checks the ``host_header``. If it's not present, then
    ``server`` is used. The host will only contain the port if it is
    different than the standard port for the protocol.

    Optionally, verify that the host is trusted using
    :func:`host_is_trusted` and raise a
    :exc:`~werkzeug.exceptions.SecurityError` if it is not.

    :param scheme: The protocol the request used, like ``"https"``.
    :param host_header: The ``Host`` header value.
    :param server: Address of the server. ``(host, port)``, or
        ``(path, None)`` for unix sockets.
    :param trusted_hosts: A list of trusted host names.

    :return: Host, with port if necessary.
    :raise ~werkzeug.exceptions.SecurityError: If the host is not
        trusted.

    .. versionchanged:: 3.1.3
        If ``SERVER_NAME`` is IPv6, it is wrapped in ``[]``.
    """
    host = ""

    if host_header is not None:
        host = host_header
    elif server is not None:
        host = server[0]

        # If SERVER_NAME is IPv6, wrap it in [] to match Host header.
        # Check for : because domain or IPv4 can't have that.
        if ":" in host and host[0] != "[":
            host = f"[{host}]"

        if server[1] is not None:
            host = f"{host}:{server[1]}"

    if scheme in {"http", "ws"} and host.endswith(":80"):
        host = host[:-3]
    elif scheme in {"https", "wss"} and host.endswith(":443"):
        host = host[:-4]

    if trusted_hosts is not None:
        if not host_is_trusted(host, trusted_hosts):
            raise SecurityError(f"Host {host!r} is not trusted.")

    return host


def get_current_url(
    scheme: str,
    host: str,
    root_path: str | None = None,
    path: str | None = None,
    query_string: bytes | None = None,
) -> str:
    """Recreate the URL for a request. If an optional part isn't
    provided, it and subsequent parts are not included in the URL.

    The URL is an IRI, not a URI, so it may contain Unicode characters.
    Use :func:`~werkzeug.urls.iri_to_uri` to convert it to ASCII.

    :param scheme: The protocol the request used, like ``"https"``.
    :param host: The host the request was made to. See :func:`get_host`.
    :param root_path: Prefix that the application is mounted under. This
        is prepended to ``path``.
    :param path: The path part of the URL after ``root_path``.
    :param query_string: The portion of the URL after the "?".
    """
    url = [scheme, "://", host]

    if root_path is None:
        url.append("/")
        return uri_to_iri("".join(url))

    # safe = https://url.spec.whatwg.org/#url-path-segment-string
    # as well as percent for things that are already quoted
    url.append(quote(root_path.rstrip("/"), safe="!$&'()*+,/:;=@%"))
    url.append("/")

    if path is None:
        return uri_to_iri("".join(url))

    url.append(quote(path.lstrip("/"), safe="!$&'()*+,/:;=@%"))

    if query_string:
        url.append("?")
        url.append(quote(query_string, safe="!$&'()*+,/:;=?@%"))

    return uri_to_iri("".join(url))


def get_content_length(
    http_content_length: str | None = None,
    http_transfer_encoding: str | None = None,
) -> int | None:
    """Return the ``Content-Length`` header value as an int. If the header is not given
    or the ``Transfer-Encoding`` header is ``chunked``, ``None`` is returned to indicate
    a streaming request. If the value is not an integer, or negative, 0 is returned.

    :param http_content_length: The Content-Length HTTP header.
    :param http_transfer_encoding: The Transfer-Encoding HTTP header.

    .. versionadded:: 2.2
    """
    if http_transfer_encoding == "chunked" or http_content_length is None:
        return None

    try:
        return max(0, _plain_int(http_content_length))
    except ValueError:
        return 0

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\jinja2\exceptions.py ===
import typing as t

if t.TYPE_CHECKING:
    from .runtime import Undefined


class TemplateError(Exception):
    """Baseclass for all template errors."""

    def __init__(self, message: t.Optional[str] = None) -> None:
        super().__init__(message)

    @property
    def message(self) -> t.Optional[str]:
        return self.args[0] if self.args else None


class TemplateNotFound(IOError, LookupError, TemplateError):
    """Raised if a template does not exist.

    .. versionchanged:: 2.11
        If the given name is :class:`Undefined` and no message was
        provided, an :exc:`UndefinedError` is raised.
    """

    # Silence the Python warning about message being deprecated since
    # it's not valid here.
    message: t.Optional[str] = None

    def __init__(
        self,
        name: t.Optional[t.Union[str, "Undefined"]],
        message: t.Optional[str] = None,
    ) -> None:
        IOError.__init__(self, name)

        if message is None:
            from .runtime import Undefined

            if isinstance(name, Undefined):
                name._fail_with_undefined_error()

            message = name

        self.message = message
        self.name = name
        self.templates = [name]

    def __str__(self) -> str:
        return str(self.message)


class TemplatesNotFound(TemplateNotFound):
    """Like :class:`TemplateNotFound` but raised if multiple templates
    are selected.  This is a subclass of :class:`TemplateNotFound`
    exception, so just catching the base exception will catch both.

    .. versionchanged:: 2.11
        If a name in the list of names is :class:`Undefined`, a message
        about it being undefined is shown rather than the empty string.

    .. versionadded:: 2.2
    """

    def __init__(
        self,
        names: t.Sequence[t.Union[str, "Undefined"]] = (),
        message: t.Optional[str] = None,
    ) -> None:
        if message is None:
            from .runtime import Undefined

            parts = []

            for name in names:
                if isinstance(name, Undefined):
                    parts.append(name._undefined_message)
                else:
                    parts.append(name)

            parts_str = ", ".join(map(str, parts))
            message = f"none of the templates given were found: {parts_str}"

        super().__init__(names[-1] if names else None, message)
        self.templates = list(names)


class TemplateSyntaxError(TemplateError):
    """Raised to tell the user that there is a problem with the template."""

    def __init__(
        self,
        message: str,
        lineno: int,
        name: t.Optional[str] = None,
        filename: t.Optional[str] = None,
    ) -> None:
        super().__init__(message)
        self.lineno = lineno
        self.name = name
        self.filename = filename
        self.source: t.Optional[str] = None

        # this is set to True if the debug.translate_syntax_error
        # function translated the syntax error into a new traceback
        self.translated = False

    def __str__(self) -> str:
        # for translated errors we only return the message
        if self.translated:
            return t.cast(str, self.message)

        # otherwise attach some stuff
        location = f"line {self.lineno}"
        name = self.filename or self.name
        if name:
            location = f'File "{name}", {location}'
        lines = [t.cast(str, self.message), "  " + location]

        # if the source is set, add the line to the output
        if self.source is not None:
            try:
                line = self.source.splitlines()[self.lineno - 1]
            except IndexError:
                pass
            else:
                lines.append("    " + line.strip())

        return "\n".join(lines)

    def __reduce__(self):  # type: ignore
        # https://bugs.python.org/issue1692335 Exceptions that take
        # multiple required arguments have problems with pickling.
        # Without this, raises TypeError: __init__() missing 1 required
        # positional argument: 'lineno'
        return self.__class__, (self.message, self.lineno, self.name, self.filename)


class TemplateAssertionError(TemplateSyntaxError):
    """Like a template syntax error, but covers cases where something in the
    template caused an error at compile time that wasn't necessarily caused
    by a syntax error.  However it's a direct subclass of
    :exc:`TemplateSyntaxError` and has the same attributes.
    """


class TemplateRuntimeError(TemplateError):
    """A generic runtime error in the template engine.  Under some situations
    Jinja may raise this exception.
    """


class UndefinedError(TemplateRuntimeError):
    """Raised if a template tries to operate on :class:`Undefined`."""


class SecurityError(TemplateRuntimeError):
    """Raised if a template tries to do something insecure if the
    sandbox is enabled.
    """


class FilterArgumentError(TemplateRuntimeError):
    """This error is raised if a filter was called with inappropriate
    arguments
    """

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\formula\translate.py ===
"""
This module contains code to translate formulae across cells in a worksheet.

The idea is that if A1 has formula "=B1+C1", then translating it to cell A2
results in formula "=B2+C2". The algorithm relies on the formula tokenizer
to identify the parts of the formula that need to change.

"""

import re
from .tokenizer import Tokenizer, Token
from openpyxl.utils import (
    coordinate_to_tuple,
    column_index_from_string,
    get_column_letter
)

class TranslatorError(Exception):
    """
    Raised when a formula can't be translated across cells.

    This error arises when a formula's references would be translated outside
    the worksheet's bounds on the top or left. Excel represents these
    situations with a #REF! literal error. E.g., if the formula at B2 is
    '=A1', attempting to translate the formula to B1 raises TranslatorError,
    since there's no cell above A1. Similarly, translating the same formula
    from B2 to A2 raises TranslatorError, since there's no cell to the left of
    A1.

    """


class Translator:

    """
    Modifies a formula so that it can be translated from one cell to another.

    `formula`: The str string to translate. Must include the leading '='
               character.
    `origin`: The cell address (in A1 notation) where this formula was
              defined (excluding the worksheet name).

    """

    def __init__(self, formula, origin):
        # Excel errors out when a workbook has formulae in R1C1 notation,
        # regardless of the calcPr:refMode setting, so I'm assuming the
        # formulae stored in the workbook must be in A1 notation.
        self.row, self.col = coordinate_to_tuple(origin)
        self.tokenizer = Tokenizer(formula)

    def get_tokens(self):
        "Returns a list with the tokens comprising the formula."
        return self.tokenizer.items

    ROW_RANGE_RE = re.compile(r"(\$?[1-9][0-9]{0,6}):(\$?[1-9][0-9]{0,6})$")
    COL_RANGE_RE = re.compile(r"(\$?[A-Za-z]{1,3}):(\$?[A-Za-z]{1,3})$")
    CELL_REF_RE = re.compile(r"(\$?[A-Za-z]{1,3})(\$?[1-9][0-9]{0,6})$")

    @staticmethod
    def translate_row(row_str, rdelta):
        """
        Translate a range row-snippet by the given number of rows.
        """
        if row_str.startswith('$'):
            return row_str
        else:
            new_row = int(row_str) + rdelta
            if new_row <= 0:
                raise TranslatorError("Formula out of range")
            return str(new_row)

    @staticmethod
    def translate_col(col_str, cdelta):
        """
        Translate a range col-snippet by the given number of columns
        """
        if col_str.startswith('$'):
            return col_str
        else:
            try:
                return get_column_letter(
                    column_index_from_string(col_str) + cdelta)
            except ValueError:
                raise TranslatorError("Formula out of range")

    @staticmethod
    def strip_ws_name(range_str):
        "Splits out the worksheet reference, if any, from a range reference."
        # This code assumes that named ranges cannot contain any exclamation
        # marks. Excel refuses to create these (even using VBA), and
        # complains of a corrupt workbook when there are names with
        # exclamation marks. The ECMA spec only states that named ranges will
        # be of `ST_Xstring` type, which in theory allows '!' (char code
        # 0x21) per http://www.w3.org/TR/xml/#charsets
        if '!' in range_str:
            sheet, range_str = range_str.rsplit('!', 1)
            return sheet + "!", range_str
        return "", range_str

    @classmethod
    def translate_range(cls, range_str, rdelta, cdelta):
        """
        Translate an A1-style range reference to the destination cell.

        `rdelta`: the row offset to add to the range
        `cdelta`: the column offset to add to the range
        `range_str`: an A1-style reference to a range. Potentially includes
                     the worksheet reference. Could also be a named range.

        """
        ws_part, range_str = cls.strip_ws_name(range_str)
        match = cls.ROW_RANGE_RE.match(range_str)  # e.g. `3:4`
        if match is not None:
            return (ws_part + cls.translate_row(match.group(1), rdelta) + ":"
                    + cls.translate_row(match.group(2), rdelta))
        match = cls.COL_RANGE_RE.match(range_str)  # e.g. `A:BC`
        if match is not None:
            return (ws_part + cls.translate_col(match.group(1), cdelta) + ':'
                    + cls.translate_col(match.group(2), cdelta))
        if ':' in range_str: # e.g. `A1:B5`
            # The check is necessarily general because range references can
            # have one or both endpoints specified by named ranges. I.e.,
            # `named_range:C2`, `C2:named_range`, and `name1:name2` are all
            # valid references. Further, Excel allows chaining multiple
            # colons together (with unclear meaning)
            return ws_part + ":".join(
                cls.translate_range(piece, rdelta, cdelta)
                for piece in range_str.split(':'))
        match = cls.CELL_REF_RE.match(range_str)
        if match is None:  # Must be a named range
            return range_str
        return (ws_part + cls.translate_col(match.group(1), cdelta)
                + cls.translate_row(match.group(2), rdelta))

    def translate_formula(self, dest=None, row_delta=0, col_delta=0):
        """
        Convert the formula into A1 notation, or as row and column coordinates

        The formula is converted into A1 assuming it is assigned to the cell
        whose address is `dest` (no worksheet name).

        """
        tokens = self.get_tokens()
        if not tokens:
            return ""
        elif tokens[0].type == Token.LITERAL:
            return tokens[0].value
        out = ['=']
        # per the spec:
        # A compliant producer or consumer considers a defined name in the
        # range A1-XFD1048576 to be an error. All other names outside this
        # range can be defined as names and overrides a cell reference if an
        # ambiguity exists. (I.18.2.5)
        if dest:
            row, col = coordinate_to_tuple(dest)
            row_delta = row - self.row
            col_delta = col - self.col
        for token in tokens:
            if (token.type == Token.OPERAND
                and token.subtype == Token.RANGE):
                out.append(self.translate_range(token.value, row_delta,
                                                col_delta))
            else:
                out.append(token.value)
        return "".join(out)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\workbook\child.py ===
# Copyright (c) 2010-2024 openpyxl

import re
import warnings

from openpyxl.worksheet.header_footer import HeaderFooter

"""
Base class for worksheets, chartsheets, etc. that can be added to workbooks
"""

INVALID_TITLE_REGEX = re.compile(r'[\\*?:/\[\]]')


def avoid_duplicate_name(names, value):
    """
    Naive check to see whether name already exists.
    If name does exist suggest a name using an incrementer
    Duplicates are case insensitive
    """
    # Check for an absolute match in which case we need to find an alternative
    match = [n for n in names if n.lower() == value.lower()]
    if match:
        names = u",".join(names)
        sheet_title_regex = re.compile(f'(?P<title>{re.escape(value)})(?P<count>\\d*),?', re.I)
        matches = sheet_title_regex.findall(names)
        if matches:
            # use name, but append with the next highest integer
            counts = [int(idx) for (t, idx) in matches if idx.isdigit()]
            highest = 0
            if counts:
                highest = max(counts)
            value = u"{0}{1}".format(value, highest + 1)
    return value


class _WorkbookChild:

    __title = ""
    _id = None
    _path = "{0}"
    _parent = None
    _default_title = "Sheet"

    def __init__(self, parent=None, title=None):
        self._parent = parent
        self.title = title or self._default_title
        self.HeaderFooter = HeaderFooter()


    def __repr__(self):
        return '<{0} "{1}">'.format(self.__class__.__name__, self.title)


    @property
    def parent(self):
        return self._parent


    @property
    def encoding(self):
        return self._parent.encoding


    @property
    def title(self):
        return self.__title


    @title.setter
    def title(self, value):
        """
        Set a sheet title, ensuring it is valid.
        Limited to 31 characters, no special characters.
        Duplicate titles will be incremented numerically
        """
        if not self._parent:
            return

        if not value:
            raise ValueError("Title must have at least one character")

        if hasattr(value, "decode"):
            if not isinstance(value, str):
                try:
                    value = value.decode("ascii")
                except UnicodeDecodeError:
                    raise ValueError("Worksheet titles must be str")

        m = INVALID_TITLE_REGEX.search(value)
        if m:
            msg = "Invalid character {0} found in sheet title".format(m.group(0))
            raise ValueError(msg)

        if self.title is not None and self.title != value:
            value = avoid_duplicate_name(self.parent.sheetnames, value)

        if len(value) > 31:
            warnings.warn("Title is more than 31 characters. Some applications may not be able to read the file")

        self.__title = value


    @property
    def oddHeader(self):
        return self.HeaderFooter.oddHeader


    @oddHeader.setter
    def oddHeader(self, value):
        self.HeaderFooter.oddHeader = value


    @property
    def oddFooter(self):
        return self.HeaderFooter.oddFooter


    @oddFooter.setter
    def oddFooter(self, value):
        self.HeaderFooter.oddFooter = value


    @property
    def evenHeader(self):
        return self.HeaderFooter.evenHeader


    @evenHeader.setter
    def evenHeader(self, value):
        self.HeaderFooter.evenHeader = value


    @property
    def evenFooter(self):
        return self.HeaderFooter.evenFooter


    @evenFooter.setter
    def evenFooter(self, value):
        self.HeaderFooter.evenFooter = value


    @property
    def firstHeader(self):
        return self.HeaderFooter.firstHeader


    @firstHeader.setter
    def firstHeader(self, value):
        self.HeaderFooter.firstHeader = value


    @property
    def firstFooter(self):
        return self.HeaderFooter.firstFooter


    @firstFooter.setter
    def firstFooter(self, value):
        self.HeaderFooter.firstFooter = value


    @property
    def path(self):
        return self._path.format(self._id)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\discrete\recurrences.py ===
"""
Recurrences
"""

from sympy.core import S, sympify
from sympy.utilities.iterables import iterable
from sympy.utilities.misc import as_int


def linrec(coeffs, init, n):
    r"""
    Evaluation of univariate linear recurrences of homogeneous type
    having coefficients independent of the recurrence variable.

    Parameters
    ==========

    coeffs : iterable
        Coefficients of the recurrence
    init : iterable
        Initial values of the recurrence
    n : Integer
        Point of evaluation for the recurrence

    Notes
    =====

    Let `y(n)` be the recurrence of given type, ``c`` be the sequence
    of coefficients, ``b`` be the sequence of initial/base values of the
    recurrence and ``k`` (equal to ``len(c)``) be the order of recurrence.
    Then,

    .. math :: y(n) = \begin{cases} b_n & 0 \le n < k \\
        c_0 y(n-1) + c_1 y(n-2) + \cdots + c_{k-1} y(n-k) & n \ge k
        \end{cases}

    Let `x_0, x_1, \ldots, x_n` be a sequence and consider the transformation
    that maps each polynomial `f(x)` to `T(f(x))` where each power `x^i` is
    replaced by the corresponding value `x_i`. The sequence is then a solution
    of the recurrence if and only if `T(x^i p(x)) = 0` for each `i \ge 0` where
    `p(x) = x^k - c_0 x^(k-1) - \cdots - c_{k-1}` is the characteristic
    polynomial.

    Then `T(f(x)p(x)) = 0` for each polynomial `f(x)` (as it is a linear
    combination of powers `x^i`). Now, if `x^n` is congruent to
    `g(x) = a_0 x^0 + a_1 x^1 + \cdots + a_{k-1} x^{k-1}` modulo `p(x)`, then
    `T(x^n) = x_n` is equal to
    `T(g(x)) = a_0 x_0 + a_1 x_1 + \cdots + a_{k-1} x_{k-1}`.

    Computation of `x^n`,
    given `x^k = c_0 x^{k-1} + c_1 x^{k-2} + \cdots + c_{k-1}`
    is performed using exponentiation by squaring (refer to [1_]) with
    an additional reduction step performed to retain only first `k` powers
    of `x` in the representation of `x^n`.

    Examples
    ========

    >>> from sympy.discrete.recurrences import linrec
    >>> from sympy.abc import x, y, z

    >>> linrec(coeffs=[1, 1], init=[0, 1], n=10)
    55

    >>> linrec(coeffs=[1, 1], init=[x, y], n=10)
    34*x + 55*y

    >>> linrec(coeffs=[x, y], init=[0, 1], n=5)
    x**2*y + x*(x**3 + 2*x*y) + y**2

    >>> linrec(coeffs=[1, 2, 3, 0, 0, 4], init=[x, y, z], n=16)
    13576*x + 5676*y + 2356*z

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Exponentiation_by_squaring
    .. [2] https://en.wikipedia.org/w/index.php?title=Modular_exponentiation&section=6#Matrices

    See Also
    ========

    sympy.polys.agca.extensions.ExtensionElement.__pow__

    """

    if not coeffs:
        return S.Zero

    if not iterable(coeffs):
        raise TypeError("Expected a sequence of coefficients for"
                        " the recurrence")

    if not iterable(init):
        raise TypeError("Expected a sequence of values for the initialization"
                        " of the recurrence")

    n = as_int(n)
    if n < 0:
        raise ValueError("Point of evaluation of recurrence must be a "
                        "non-negative integer")

    c = [sympify(arg) for arg in coeffs]
    b = [sympify(arg) for arg in init]
    k = len(c)

    if len(b) > k:
        raise TypeError("Count of initial values should not exceed the "
                        "order of the recurrence")
    else:
        b += [S.Zero]*(k - len(b)) # remaining initial values default to zero

    if n < k:
        return b[n]
    terms = [u*v for u, v in zip(linrec_coeffs(c, n), b)]
    return sum(terms[:-1], terms[-1])


def linrec_coeffs(c, n):
    r"""
    Compute the coefficients of n'th term in linear recursion
    sequence defined by c.

    `x^k = c_0 x^{k-1} + c_1 x^{k-2} + \cdots + c_{k-1}`.

    It computes the coefficients by using binary exponentiation.
    This function is used by `linrec` and `_eval_pow_by_cayley`.

    Parameters
    ==========

    c = coefficients of the divisor polynomial
    n = exponent of x, so dividend is x^n

    """

    k = len(c)

    def _square_and_reduce(u, offset):
        # squares `(u_0 + u_1 x + u_2 x^2 + \cdots + u_{k-1} x^k)` (and
        # multiplies by `x` if offset is 1) and reduces the above result of
        # length upto `2k` to `k` using the characteristic equation of the
        # recurrence given by, `x^k = c_0 x^{k-1} + c_1 x^{k-2} + \cdots + c_{k-1}`

        w = [S.Zero]*(2*len(u) - 1 + offset)
        for i, p in enumerate(u):
            for j, q in enumerate(u):
                w[offset + i + j] += p*q

        for j in range(len(w) - 1, k - 1, -1):
            for i in range(k):
                w[j - i - 1] += w[j]*c[i]

        return w[:k]

    def _final_coeffs(n):
        # computes the final coefficient list - `cf` corresponding to the
        # point at which recurrence is to be evalauted - `n`, such that,
        # `y(n) = cf_0 y(k-1) + cf_1 y(k-2) + \cdots + cf_{k-1} y(0)`

        if n < k:
            return [S.Zero]*n + [S.One] + [S.Zero]*(k - n - 1)
        else:
            return _square_and_reduce(_final_coeffs(n // 2), n % 2)

    return _final_coeffs(n)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\quantum\anticommutator.py ===
"""The anti-commutator: ``{A,B} = A*B + B*A``."""

from sympy.core.expr import Expr
from sympy.core.kind import KindDispatcher
from sympy.core.mul import Mul
from sympy.core.numbers import Integer
from sympy.core.singleton import S
from sympy.printing.pretty.stringpict import prettyForm

from sympy.physics.quantum.dagger import Dagger
from sympy.physics.quantum.kind import _OperatorKind, OperatorKind

__all__ = [
    'AntiCommutator'
]

#-----------------------------------------------------------------------------
# Anti-commutator
#-----------------------------------------------------------------------------


class AntiCommutator(Expr):
    """The standard anticommutator, in an unevaluated state.

    Explanation
    ===========

    Evaluating an anticommutator is defined [1]_ as: ``{A, B} = A*B + B*A``.
    This class returns the anticommutator in an unevaluated form.  To evaluate
    the anticommutator, use the ``.doit()`` method.

    Canonical ordering of an anticommutator is ``{A, B}`` for ``A < B``. The
    arguments of the anticommutator are put into canonical order using
    ``__cmp__``. If ``B < A``, then ``{A, B}`` is returned as ``{B, A}``.

    Parameters
    ==========

    A : Expr
        The first argument of the anticommutator {A,B}.
    B : Expr
        The second argument of the anticommutator {A,B}.

    Examples
    ========

    >>> from sympy import symbols
    >>> from sympy.physics.quantum import AntiCommutator
    >>> from sympy.physics.quantum import Operator, Dagger
    >>> x, y = symbols('x,y')
    >>> A = Operator('A')
    >>> B = Operator('B')

    Create an anticommutator and use ``doit()`` to multiply them out.

    >>> ac = AntiCommutator(A,B); ac
    {A,B}
    >>> ac.doit()
    A*B + B*A

    The commutator orders it arguments in canonical order:

    >>> ac = AntiCommutator(B,A); ac
    {A,B}

    Commutative constants are factored out:

    >>> AntiCommutator(3*x*A,x*y*B)
    3*x**2*y*{A,B}

    Adjoint operations applied to the anticommutator are properly applied to
    the arguments:

    >>> Dagger(AntiCommutator(A,B))
    {Dagger(A),Dagger(B)}

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Commutator
    """
    is_commutative = False

    _kind_dispatcher = KindDispatcher("AntiCommutator_kind_dispatcher", commutative=True)

    @property
    def kind(self):
        arg_kinds = (a.kind for a in self.args)
        return self._kind_dispatcher(*arg_kinds)

    def __new__(cls, A, B):
        r = cls.eval(A, B)
        if r is not None:
            return r
        obj = Expr.__new__(cls, A, B)
        return obj

    @classmethod
    def eval(cls, a, b):
        if not (a and b):
            return S.Zero
        if a == b:
            return Integer(2)*a**2
        if a.is_commutative or b.is_commutative:
            return Integer(2)*a*b

        # [xA,yB]  ->  xy*[A,B]
        ca, nca = a.args_cnc()
        cb, ncb = b.args_cnc()
        c_part = ca + cb
        if c_part:
            return Mul(Mul(*c_part), cls(Mul._from_args(nca), Mul._from_args(ncb)))

        # Canonical ordering of arguments
        #The Commutator [A,B] is on canonical form if A < B.
        if a.compare(b) == 1:
            return cls(b, a)

    def doit(self, **hints):
        """ Evaluate anticommutator """
        # Keep the import of Operator here to avoid problems with
        # circular imports.
        from sympy.physics.quantum.operator import Operator
        A = self.args[0]
        B = self.args[1]
        if isinstance(A, Operator) and isinstance(B, Operator):
            try:
                comm = A._eval_anticommutator(B, **hints)
            except NotImplementedError:
                try:
                    comm = B._eval_anticommutator(A, **hints)
                except NotImplementedError:
                    comm = None
            if comm is not None:
                return comm.doit(**hints)
        return (A*B + B*A).doit(**hints)

    def _eval_adjoint(self):
        return AntiCommutator(Dagger(self.args[0]), Dagger(self.args[1]))

    def _sympyrepr(self, printer, *args):
        return "%s(%s,%s)" % (
            self.__class__.__name__, printer._print(
                self.args[0]), printer._print(self.args[1])
        )

    def _sympystr(self, printer, *args):
        return "{%s,%s}" % (
            printer._print(self.args[0]), printer._print(self.args[1]))

    def _pretty(self, printer, *args):
        pform = printer._print(self.args[0], *args)
        pform = prettyForm(*pform.right(prettyForm(',')))
        pform = prettyForm(*pform.right(printer._print(self.args[1], *args)))
        pform = prettyForm(*pform.parens(left='{', right='}'))
        return pform

    def _latex(self, printer, *args):
        return "\\left\\{%s,%s\\right\\}" % tuple([
            printer._print(arg, *args) for arg in self.args])


@AntiCommutator._kind_dispatcher.register(_OperatorKind, _OperatorKind)
def find_op_kind(e1, e2):
    """Find the kind of an anticommutator of two OperatorKinds."""
    return OperatorKind

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\werkzeug\security.py ===
from __future__ import annotations

import hashlib
import hmac
import os
import posixpath
import secrets

SALT_CHARS = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
DEFAULT_PBKDF2_ITERATIONS = 1_000_000

_os_alt_seps: list[str] = list(
    sep for sep in [os.sep, os.path.altsep] if sep is not None and sep != "/"
)


def gen_salt(length: int) -> str:
    """Generate a random string of SALT_CHARS with specified ``length``."""
    if length <= 0:
        raise ValueError("Salt length must be at least 1.")

    return "".join(secrets.choice(SALT_CHARS) for _ in range(length))


def _hash_internal(method: str, salt: str, password: str) -> tuple[str, str]:
    method, *args = method.split(":")
    salt_bytes = salt.encode()
    password_bytes = password.encode()

    if method == "scrypt":
        if not args:
            n = 2**15
            r = 8
            p = 1
        else:
            try:
                n, r, p = map(int, args)
            except ValueError:
                raise ValueError("'scrypt' takes 3 arguments.") from None

        maxmem = 132 * n * r * p  # ideally 128, but some extra seems needed
        return (
            hashlib.scrypt(
                password_bytes, salt=salt_bytes, n=n, r=r, p=p, maxmem=maxmem
            ).hex(),
            f"scrypt:{n}:{r}:{p}",
        )
    elif method == "pbkdf2":
        len_args = len(args)

        if len_args == 0:
            hash_name = "sha256"
            iterations = DEFAULT_PBKDF2_ITERATIONS
        elif len_args == 1:
            hash_name = args[0]
            iterations = DEFAULT_PBKDF2_ITERATIONS
        elif len_args == 2:
            hash_name = args[0]
            iterations = int(args[1])
        else:
            raise ValueError("'pbkdf2' takes 2 arguments.")

        return (
            hashlib.pbkdf2_hmac(
                hash_name, password_bytes, salt_bytes, iterations
            ).hex(),
            f"pbkdf2:{hash_name}:{iterations}",
        )
    else:
        raise ValueError(f"Invalid hash method '{method}'.")


def generate_password_hash(
    password: str, method: str = "scrypt", salt_length: int = 16
) -> str:
    """Securely hash a password for storage. A password can be compared to a stored hash
    using :func:`check_password_hash`.

    The following methods are supported:

    -   ``scrypt``, the default. The parameters are ``n``, ``r``, and ``p``, the default
        is ``scrypt:32768:8:1``. See :func:`hashlib.scrypt`.
    -   ``pbkdf2``, less secure. The parameters are ``hash_method`` and ``iterations``,
        the default is ``pbkdf2:sha256:600000``. See :func:`hashlib.pbkdf2_hmac`.

    Default parameters may be updated to reflect current guidelines, and methods may be
    deprecated and removed if they are no longer considered secure. To migrate old
    hashes, you may generate a new hash when checking an old hash, or you may contact
    users with a link to reset their password.

    :param password: The plaintext password.
    :param method: The key derivation function and parameters.
    :param salt_length: The number of characters to generate for the salt.

    .. versionchanged:: 3.1
        The default iterations for pbkdf2 was increased to 1,000,000.

    .. versionchanged:: 2.3
        Scrypt support was added.

    .. versionchanged:: 2.3
        The default iterations for pbkdf2 was increased to 600,000.

    .. versionchanged:: 2.3
        All plain hashes are deprecated and will not be supported in Werkzeug 3.0.
    """
    salt = gen_salt(salt_length)
    h, actual_method = _hash_internal(method, salt, password)
    return f"{actual_method}${salt}${h}"


def check_password_hash(pwhash: str, password: str) -> bool:
    """Securely check that the given stored password hash, previously generated using
    :func:`generate_password_hash`, matches the given password.

    Methods may be deprecated and removed if they are no longer considered secure. To
    migrate old hashes, you may generate a new hash when checking an old hash, or you
    may contact users with a link to reset their password.

    :param pwhash: The hashed password.
    :param password: The plaintext password.

    .. versionchanged:: 2.3
        All plain hashes are deprecated and will not be supported in Werkzeug 3.0.
    """
    try:
        method, salt, hashval = pwhash.split("$", 2)
    except ValueError:
        return False

    return hmac.compare_digest(_hash_internal(method, salt, password)[0], hashval)


def safe_join(directory: str, *pathnames: str) -> str | None:
    """Safely join zero or more untrusted path components to a base
    directory to avoid escaping the base directory.

    :param directory: The trusted base directory.
    :param pathnames: The untrusted path components relative to the
        base directory.
    :return: A safe path, otherwise ``None``.
    """
    if not directory:
        # Ensure we end up with ./path if directory="" is given,
        # otherwise the first untrusted part could become trusted.
        directory = "."

    parts = [directory]

    for filename in pathnames:
        if filename != "":
            filename = posixpath.normpath(filename)

        if (
            any(sep in filename for sep in _os_alt_seps)
            or os.path.isabs(filename)
            # ntpath.isabs doesn't catch this on Python < 3.11
            or filename.startswith("/")
            or filename == ".."
            or filename.startswith("../")
        ):
            return None

        parts.append(filename)

    return posixpath.join(*parts)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_simplified_layernorm.py ===
import logging

from fusion_base import Fusion
from fusion_skiplayernorm import FusionSkipLayerNormalization
from onnx import helper
from onnx_model import OnnxModel

logger = logging.getLogger(__name__)


class FusionSimplifiedLayerNormalization(Fusion):
    def __init__(self, model: OnnxModel):
        super().__init__(model, "SimplifiedLayerNormalization", "Mul")

    def fuse(self, node, input_name_to_nodes: dict, output_name_to_node: dict):
        if node.op_type != "Mul":
            return

        sim_ln_nodes = None
        # RMSNorm formula:
        #   S = Pow(X, 2) or S = Mul(X, X)
        #   MS = ReduceMean(S)
        #   MSEps = Add(MS, epsilon)
        #   RMS = Sqrt(MSEps)
        #   InvRMS = Div(1, RMS) or InvRMS = Reciprocal(RMS)
        #   Normalized = Mul(D, InvRMS)
        #   Y = Mul(Normalized, Scale)
        #
        #  (root_input) ----------------------------------------+
        #       |                                               |
        #       v                                               v
        #      Pow --> ReduceMean --> Add ---> Sqrt --> Div --> Mul --> Mul (node)
        #      (B=2)                  (A/B=eps)         (A=1)           (A/B=scale)
        #
        #  (root_input) ----------------------------------------+
        #      | |                                              |
        #      v v                                              v
        #      Mul --> ReduceMean --> Add ---> Sqrt --> Div --> Mul --> Mul (node)
        #      (B=2)                  (A/B=eps)         (A=1)           (A/B=scale)
        #
        return_indice = []
        sim_ln_nodes = self.model.match_parent_path(
            node,
            ["Mul", "Div", "Sqrt", "Add", "ReduceMean"],
            [None, 1, 1, 0, None],
            output_name_to_node=output_name_to_node,
            return_indice=return_indice,
        )

        if sim_ln_nodes:
            mul_node, div_node, _sqrt_node, add_node, reduce_mean_node = sim_ln_nodes
            if not self.model.has_constant_input(div_node, 1.0):
                return
            node_parent = mul_node
        else:
            # Div(1, RMS) can also be represented as Reciprocal(RMS) like
            #
            #  (root_input) -----------------------------------------------+
            #       |                                                      |
            #       v                                                      v
            #      Pow --> ReduceMean --> Add ---> Sqrt --> Reciprocal --> Mul --> Mul (node)
            #      (B=2)                  (A/B=eps)                                (A/B=scale)
            #
            #  (root_input) -----------------------------------------------+
            #      | |                                                     |
            #      v v                                                     v
            #      Mul --> ReduceMean --> Add ---> Sqrt --> Reciprocal --> Mul --> Mul (node)
            #      (B=2)                  (A/B=eps)                                (A/B=scale)
            #
            return_indice = []
            sim_ln_nodes = self.model.match_parent_path(
                node,
                ["Mul", "Reciprocal", "Sqrt", "Add", "ReduceMean"],
                [None, 1, 0, 0, None],
                output_name_to_node=output_name_to_node,
                return_indice=return_indice,
            )
            if sim_ln_nodes is not None:
                mul_node, _reciprocal_node, _sqrt_node, add_node, reduce_mean_node = sim_ln_nodes
                node_parent = mul_node
            else:
                #  (root_input) --------------------------------+
                #       |                                       |
                #       v                                       v
                #      Pow --> ReduceMean --> Add ---> Sqrt --> Div --> Mul (node)
                #      (B=2)                  (A/B=eps)                 (A/B=scale)
                #
                #  (root_input) --------------------------------+
                #      | |                                      |
                #      v v                                      v
                #      Mul --> ReduceMean --> Add ---> Sqrt --> Div --> Mul (node)
                #      (B=2)                  (A/B=eps)                 (A/B=scale)
                #
                return_indice = []
                sim_ln_nodes = self.model.match_parent_path(
                    node,
                    ["Div", "Sqrt", "Add", "ReduceMean"],
                    [None, 1, 0, None],
                    output_name_to_node=output_name_to_node,
                    return_indice=return_indice,
                )
                if sim_ln_nodes is not None:
                    div_node, _sqrt_node, add_node, reduce_mean_node = sim_ln_nodes
                    node_parent = div_node
                else:
                    return

        reduce_mean_parent = self.model.get_parent(reduce_mean_node, 0, output_name_to_node)
        if reduce_mean_parent is None or reduce_mean_parent.op_type not in ["Pow", "Mul"]:
            return

        if reduce_mean_parent.op_type == "Pow":
            if self.model.find_constant_input(reduce_mean_parent, 2.0) != 1:
                return
        else:
            assert reduce_mean_parent.op_type == "Mul"
            if reduce_mean_parent[0] != reduce_mean_parent[1]:
                return

        root_input = reduce_mean_parent.input[0]
        if root_input not in node_parent.input:
            return

        _i, epsilon = self.model.get_constant_input(add_node)
        if epsilon is None or epsilon <= 0 or epsilon > 1.0e-4:
            logger.warning(f"epsilon value is not expected: {epsilon}")
            return

        # ReduceMean must have keepdims == 1
        keepdims = self.model.get_node_attribute(reduce_mean_node, "keepdims")
        if not keepdims:
            return

        # ReduceMean axes must refer only to the last dimension.
        # Axes became an input in opset 18. Before then, axes was an attribute.
        axes = self.model.get_node_attribute(reduce_mean_node, "axes")
        if (not axes) and len(reduce_mean_node.input) > 1:
            axes = self.model.get_constant_value(reduce_mean_node.input[1])
        # Make sure only one axis as required by SimplifiedLayerNormalization spec.
        if not axes or len(axes) != 1:
            return

        self.nodes_to_remove.extend(sim_ln_nodes)
        self.nodes_to_remove.append(reduce_mean_parent)
        self.nodes_to_remove.append(node)

        normalize_node = helper.make_node(
            "SimplifiedLayerNormalization",
            inputs=[root_input, node.input[1 - return_indice[0]]],
            outputs=[node.output[0]],
            name=self.model.create_node_name("SimplifiedLayerNormalization", name_prefix="RMSNorm"),
        )
        normalize_node.attribute.extend([helper.make_attribute("epsilon", float(epsilon))])
        normalize_node.attribute.extend([helper.make_attribute("axis", axes[0])])
        normalize_node.attribute.extend([helper.make_attribute("stash_type", 1)])
        self.nodes_to_add.append(normalize_node)
        self.node_name_to_graph_name[normalize_node.name] = self.this_graph_name


class FusionSkipSimplifiedLayerNormalization(FusionSkipLayerNormalization):
    def __init__(self, model: OnnxModel):
        super().__init__(model, "SkipSimplifiedLayerNormalization", "SimplifiedLayerNormalization")

    def fuse(self, node, input_name_to_nodes, output_name_to_node):
        super().fuse(node, input_name_to_nodes, output_name_to_node)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_core\_mock_clock.py ===
import time
from math import inf

from .. import _core
from .._abc import Clock
from .._util import final
from ._run import GLOBAL_RUN_CONTEXT

################################################################
# The glorious MockClock
################################################################


# Prior art:
#   https://twistedmatrix.com/documents/current/api/twisted.internet.task.Clock.html
#   https://github.com/ztellman/manifold/issues/57
@final
class MockClock(Clock):
    """A user-controllable clock suitable for writing tests.

    Args:
      rate (float): the initial :attr:`rate`.
      autojump_threshold (float): the initial :attr:`autojump_threshold`.

    .. attribute:: rate

       How many seconds of clock time pass per second of real time. Default is
       0.0, i.e. the clock only advances through manuals calls to :meth:`jump`
       or when the :attr:`autojump_threshold` is triggered. You can assign to
       this attribute to change it.

    .. attribute:: autojump_threshold

       The clock keeps an eye on the run loop, and if at any point it detects
       that all tasks have been blocked for this many real seconds (i.e.,
       according to the actual clock, not this clock), then the clock
       automatically jumps ahead to the run loop's next scheduled
       timeout. Default is :data:`math.inf`, i.e., to never autojump. You can
       assign to this attribute to change it.

       Basically the idea is that if you have code or tests that use sleeps
       and timeouts, you can use this to make it run much faster, totally
       automatically. (At least, as long as those sleeps/timeouts are
       happening inside Trio; if your test involves talking to external
       service and waiting for it to timeout then obviously we can't help you
       there.)

       You should set this to the smallest value that lets you reliably avoid
       "false alarms" where some I/O is in flight (e.g. between two halves of
       a socketpair) but the threshold gets triggered and time gets advanced
       anyway. This will depend on the details of your tests and test
       environment. If you aren't doing any I/O (like in our sleeping example
       above) then just set it to zero, and the clock will jump whenever all
       tasks are blocked.

       .. note:: If you use ``autojump_threshold`` and
          `wait_all_tasks_blocked` at the same time, then you might wonder how
          they interact, since they both cause things to happen after the run
          loop goes idle for some time. The answer is:
          `wait_all_tasks_blocked` takes priority. If there's a task blocked
          in `wait_all_tasks_blocked`, then the autojump feature treats that
          as active task and does *not* jump the clock.

    """

    def __init__(self, rate: float = 0.0, autojump_threshold: float = inf) -> None:
        # when the real clock said 'real_base', the virtual time was
        # 'virtual_base', and since then it's advanced at 'rate' virtual
        # seconds per real second.
        self._real_base = 0.0
        self._virtual_base = 0.0
        self._rate = 0.0

        # kept as an attribute so that our tests can monkeypatch it
        self._real_clock = time.perf_counter

        # use the property update logic to set initial values
        self.rate = rate
        self.autojump_threshold = autojump_threshold

    def __repr__(self) -> str:
        return f"<MockClock, time={self.current_time():.7f}, rate={self._rate} @ {id(self):#x}>"

    @property
    def rate(self) -> float:
        return self._rate

    @rate.setter
    def rate(self, new_rate: float) -> None:
        if new_rate < 0:
            raise ValueError("rate must be >= 0")
        else:
            real = self._real_clock()
            virtual = self._real_to_virtual(real)
            self._virtual_base = virtual
            self._real_base = real
            self._rate = float(new_rate)

    @property
    def autojump_threshold(self) -> float:
        return self._autojump_threshold

    @autojump_threshold.setter
    def autojump_threshold(self, new_autojump_threshold: float) -> None:
        self._autojump_threshold = float(new_autojump_threshold)
        self._try_resync_autojump_threshold()

    # runner.clock_autojump_threshold is an internal API that isn't easily
    # usable by custom third-party Clock objects. If you need access to this
    # functionality, let us know, and we'll figure out how to make a public
    # API. Discussion:
    #
    #     https://github.com/python-trio/trio/issues/1587
    def _try_resync_autojump_threshold(self) -> None:
        try:
            runner = GLOBAL_RUN_CONTEXT.runner
            if runner.is_guest:
                runner.force_guest_tick_asap()
        except AttributeError:
            pass
        else:
            if runner.clock is self:
                runner.clock_autojump_threshold = self._autojump_threshold

    # Invoked by the run loop when runner.clock_autojump_threshold is
    # exceeded.
    def _autojump(self) -> None:
        statistics = _core.current_statistics()
        jump = statistics.seconds_to_next_deadline
        if 0 < jump < inf:
            self.jump(jump)

    def _real_to_virtual(self, real: float) -> float:
        real_offset = real - self._real_base
        virtual_offset = self._rate * real_offset
        return self._virtual_base + virtual_offset

    def start_clock(self) -> None:
        self._try_resync_autojump_threshold()

    def current_time(self) -> float:
        return self._real_to_virtual(self._real_clock())

    def deadline_to_sleep_time(self, deadline: float) -> float:
        virtual_timeout = deadline - self.current_time()
        if virtual_timeout <= 0:
            return 0
        elif self._rate > 0:
            return virtual_timeout / self._rate
        else:
            return 999999999

    def jump(self, seconds: float) -> None:
        """Manually advance the clock by the given number of seconds.

        Args:
          seconds (float): the number of seconds to jump the clock forward.

        Raises:
          ValueError: if you try to pass a negative value for ``seconds``.

        """
        if seconds < 0:
            raise ValueError("time can't go backwards")
        self._virtual_base += seconds

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\urllib3\_base_connection.py ===
from __future__ import annotations

import typing

from .util.connection import _TYPE_SOCKET_OPTIONS
from .util.timeout import _DEFAULT_TIMEOUT, _TYPE_TIMEOUT
from .util.url import Url

_TYPE_BODY = typing.Union[bytes, typing.IO[typing.Any], typing.Iterable[bytes], str]


class ProxyConfig(typing.NamedTuple):
    ssl_context: ssl.SSLContext | None
    use_forwarding_for_https: bool
    assert_hostname: None | str | typing.Literal[False]
    assert_fingerprint: str | None


class _ResponseOptions(typing.NamedTuple):
    # TODO: Remove this in favor of a better
    # HTTP request/response lifecycle tracking.
    request_method: str
    request_url: str
    preload_content: bool
    decode_content: bool
    enforce_content_length: bool


if typing.TYPE_CHECKING:
    import ssl
    from typing import Protocol

    from .response import BaseHTTPResponse

    class BaseHTTPConnection(Protocol):
        default_port: typing.ClassVar[int]
        default_socket_options: typing.ClassVar[_TYPE_SOCKET_OPTIONS]

        host: str
        port: int
        timeout: None | (
            float
        )  # Instance doesn't store _DEFAULT_TIMEOUT, must be resolved.
        blocksize: int
        source_address: tuple[str, int] | None
        socket_options: _TYPE_SOCKET_OPTIONS | None

        proxy: Url | None
        proxy_config: ProxyConfig | None

        is_verified: bool
        proxy_is_verified: bool | None

        def __init__(
            self,
            host: str,
            port: int | None = None,
            *,
            timeout: _TYPE_TIMEOUT = _DEFAULT_TIMEOUT,
            source_address: tuple[str, int] | None = None,
            blocksize: int = 8192,
            socket_options: _TYPE_SOCKET_OPTIONS | None = ...,
            proxy: Url | None = None,
            proxy_config: ProxyConfig | None = None,
        ) -> None: ...

        def set_tunnel(
            self,
            host: str,
            port: int | None = None,
            headers: typing.Mapping[str, str] | None = None,
            scheme: str = "http",
        ) -> None: ...

        def connect(self) -> None: ...

        def request(
            self,
            method: str,
            url: str,
            body: _TYPE_BODY | None = None,
            headers: typing.Mapping[str, str] | None = None,
            # We know *at least* botocore is depending on the order of the
            # first 3 parameters so to be safe we only mark the later ones
            # as keyword-only to ensure we have space to extend.
            *,
            chunked: bool = False,
            preload_content: bool = True,
            decode_content: bool = True,
            enforce_content_length: bool = True,
        ) -> None: ...

        def getresponse(self) -> BaseHTTPResponse: ...

        def close(self) -> None: ...

        @property
        def is_closed(self) -> bool:
            """Whether the connection either is brand new or has been previously closed.
            If this property is True then both ``is_connected`` and ``has_connected_to_proxy``
            properties must be False.
            """

        @property
        def is_connected(self) -> bool:
            """Whether the connection is actively connected to any origin (proxy or target)"""

        @property
        def has_connected_to_proxy(self) -> bool:
            """Whether the connection has successfully connected to its proxy.
            This returns False if no proxy is in use. Used to determine whether
            errors are coming from the proxy layer or from tunnelling to the target origin.
            """

    class BaseHTTPSConnection(BaseHTTPConnection, Protocol):
        default_port: typing.ClassVar[int]
        default_socket_options: typing.ClassVar[_TYPE_SOCKET_OPTIONS]

        # Certificate verification methods
        cert_reqs: int | str | None
        assert_hostname: None | str | typing.Literal[False]
        assert_fingerprint: str | None
        ssl_context: ssl.SSLContext | None

        # Trusted CAs
        ca_certs: str | None
        ca_cert_dir: str | None
        ca_cert_data: None | str | bytes

        # TLS version
        ssl_minimum_version: int | None
        ssl_maximum_version: int | None
        ssl_version: int | str | None  # Deprecated

        # Client certificates
        cert_file: str | None
        key_file: str | None
        key_password: str | None

        def __init__(
            self,
            host: str,
            port: int | None = None,
            *,
            timeout: _TYPE_TIMEOUT = _DEFAULT_TIMEOUT,
            source_address: tuple[str, int] | None = None,
            blocksize: int = 16384,
            socket_options: _TYPE_SOCKET_OPTIONS | None = ...,
            proxy: Url | None = None,
            proxy_config: ProxyConfig | None = None,
            cert_reqs: int | str | None = None,
            assert_hostname: None | str | typing.Literal[False] = None,
            assert_fingerprint: str | None = None,
            server_hostname: str | None = None,
            ssl_context: ssl.SSLContext | None = None,
            ca_certs: str | None = None,
            ca_cert_dir: str | None = None,
            ca_cert_data: None | str | bytes = None,
            ssl_minimum_version: int | None = None,
            ssl_maximum_version: int | None = None,
            ssl_version: int | str | None = None,  # Deprecated
            cert_file: str | None = None,
            key_file: str | None = None,
            key_password: str | None = None,
        ) -> None: ...

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\tools\onnxruntime_test.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation. All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
from __future__ import annotations

import argparse
import os
import sys
from timeit import default_timer as timer

import numpy as np

import onnxruntime as onnxrt

float_dict = {
    "tensor(float16)": "float16",
    "tensor(float)": "float32",
    "tensor(double)": "float64",
}

integer_dict = {
    "tensor(int32)": "int32",
    "tensor(int8)": "int8",
    "tensor(uint8)": "uint8",
    "tensor(int16)": "int16",
    "tensor(uint16)": "uint16",
    "tensor(int64)": "int64",
    "tensor(uint64)": "uint64",
}


def generate_feeds(sess, symbolic_dims: dict | None = None):
    feeds = {}
    symbolic_dims = symbolic_dims or {}
    for input_meta in sess.get_inputs():
        # replace any symbolic dimensions
        shape = []
        for dim in input_meta.shape:
            if not dim:
                # unknown dim
                shape.append(1)
            elif isinstance(dim, str):
                # symbolic dim. see if we have a value otherwise use 1
                if dim in symbolic_dims:
                    shape.append(int(symbolic_dims[dim]))
                else:
                    shape.append(1)
            else:
                shape.append(dim)

        if input_meta.type in float_dict:
            feeds[input_meta.name] = np.random.rand(*shape).astype(float_dict[input_meta.type])
        elif input_meta.type in integer_dict:
            feeds[input_meta.name] = np.random.uniform(high=1000, size=tuple(shape)).astype(
                integer_dict[input_meta.type]
            )
        elif input_meta.type == "tensor(bool)":
            feeds[input_meta.name] = np.random.randint(2, size=tuple(shape)).astype("bool")
        else:
            print(f"unsupported input type {input_meta.type} for input {input_meta.name}")
            sys.exit(-1)
    return feeds


# simple test program for loading onnx model, feeding all inputs and running the model num_iters times.
def run_model(
    model_path,
    num_iters=1,
    debug=None,
    profile=None,
    symbolic_dims=None,
    feeds=None,
    override_initializers=True,
):
    symbolic_dims = symbolic_dims or {}
    if debug:
        print(f"Pausing execution ready for debugger to attach to pid: {os.getpid()}")
        print("Press key to continue.")
        sys.stdin.read(1)

    sess_options = None
    if profile:
        sess_options = onnxrt.SessionOptions()
        sess_options.enable_profiling = True
        sess_options.profile_file_prefix = os.path.basename(model_path)

    sess = onnxrt.InferenceSession(
        model_path,
        sess_options=sess_options,
        providers=onnxrt.get_available_providers(),
    )
    meta = sess.get_modelmeta()

    if not feeds:
        feeds = generate_feeds(sess, symbolic_dims)

    if override_initializers:
        # Starting with IR4 some initializers provide default values
        # and can be overridden (available in IR4). For IR < 4 models
        # the list would be empty
        for initializer in sess.get_overridable_initializers():
            shape = [dim if dim else 1 for dim in initializer.shape]
            if initializer.type in float_dict:
                feeds[initializer.name] = np.random.rand(*shape).astype(float_dict[initializer.type])
            elif initializer.type in integer_dict:
                feeds[initializer.name] = np.random.uniform(high=1000, size=tuple(shape)).astype(
                    integer_dict[initializer.type]
                )
            elif initializer.type == "tensor(bool)":
                feeds[initializer.name] = np.random.randint(2, size=tuple(shape)).astype("bool")
            else:
                print(f"unsupported initializer type {initializer.type} for initializer {initializer.name}")
                sys.exit(-1)

    start = timer()
    for _i in range(num_iters):
        outputs = sess.run([], feeds)  # fetch all outputs
    end = timer()

    print(f"model: {meta.graph_name}")
    print(f"version: {meta.version}")
    print(f"iterations: {num_iters}")
    print(f"avg latency: {((end - start) * 1000) / num_iters} ms")

    if profile:
        trace_file = sess.end_profiling()
        print(f"trace file written to: {trace_file}")

    return 0, feeds, num_iters > 0 and outputs


def main():
    parser = argparse.ArgumentParser(description="Simple ONNX Runtime Test Tool.")
    parser.add_argument("model_path", help="model path")
    parser.add_argument(
        "num_iters",
        nargs="?",
        type=int,
        default=1000,
        help="model run iterations. default=1000",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="pause execution to allow attaching a debugger.",
    )
    parser.add_argument("--profile", action="store_true", help="enable chrome timeline trace profiling.")
    parser.add_argument(
        "--symbolic_dims",
        default={},
        type=lambda s: dict(x.split("=") for x in s.split(",")),
        help="Comma separated name=value pairs for any symbolic dimensions in the model input. "
        "e.g. --symbolic_dims batch=1,seqlen=5. "
        "If not provided, the value of 1 will be used for all symbolic dimensions.",
    )

    args = parser.parse_args()
    exit_code, _, _ = run_model(args.model_path, args.num_iters, args.debug, args.profile, args.symbolic_dims)
    sys.exit(exit_code)


if __name__ == "__main__":
    main()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\whisper\whisper_encoder.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.  See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------

import logging
import os
import tempfile
from pathlib import Path

import numpy as np
import onnx
import torch
from float16 import convert_float_to_float16
from onnx import ModelProto
from onnx_model import OnnxModel
from transformers import WhisperConfig
from whisper_inputs import get_model_dynamic_axes, get_sample_encoder_inputs

from onnxruntime import InferenceSession

logger = logging.getLogger(__name__)


class WhisperEncoder(torch.nn.Module):
    """Whisper encoder component"""

    def __init__(self, config: WhisperConfig, model: torch.nn.Module, model_impl: str):
        super().__init__()
        self.config = config
        self.device = model.device
        self.model_impl = model_impl

        self.encoder = model.encoder if model_impl == "openai" else model.model.encoder

    def forward(self, audio_features: torch.Tensor):
        outputs = self.encoder(audio_features)
        return outputs if self.model_impl == "openai" else outputs.last_hidden_state

    def input_names(self):
        input_names = ["audio_features"]
        return input_names

    def output_names(self):
        output_names = ["encoder_hidden_states"]
        return output_names

    def dynamic_axes(self, input_names, output_names):
        dynamic_axes = get_model_dynamic_axes(self.config, input_names, output_names)
        return dynamic_axes

    def fix_layernorm_weights(self, model: ModelProto, use_fp16_inputs: bool):
        if self.model_impl == "openai" and use_fp16_inputs:
            # Cast ONNX model to float16 to ensure LayerNorm weights are converted from
            # float32 to float16 since exported model already has float16 weights everywhere
            # except for LayerNorm ops. This happens because OpenAI always upcasts to float32
            # when computing LayerNorm.
            #
            # Reference:
            # https://github.com/openai/whisper/blob/90db0de1896c23cbfaf0c58bc2d30665f709f170/whisper/model.py#L41
            model = convert_float_to_float16(model)
        return model

    def export_onnx(
        self,
        onnx_model_path: str,
        provider: str,
        verbose: bool = True,
        use_external_data_format: bool = False,
        use_fp16_inputs: bool = False,
    ):
        """Export encoder to ONNX

        Args:
            onnx_model_path (str): path to save ONNX model
            provider (str): provider to use for verifying parity on ONNX model
            verbose (bool, optional): print verbose information. Defaults to True.
            use_external_data_format (bool, optional): use external data format or not. Defaults to False.
            use_fp16_inputs (bool, optional): use float16 inputs for the audio_features. Defaults to False.
        """
        # Shape of encoder's tensors:
        # Inputs:
        #    audio_features: (batch_size, num_mels, num_frames)
        # Outputs:
        #    encoder_hidden_states: (batch_size, num_frames // 2, hidden_size)

        inputs = get_sample_encoder_inputs(
            self.config,
            self.device,
            batch_size=2,
            use_fp16=use_fp16_inputs,
        )

        input_names = self.input_names()
        output_names = self.output_names()
        dynamic_axes = self.dynamic_axes(input_names, output_names)

        Path(onnx_model_path).parent.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory() as tmp_dir_name:
            temp_onnx_model_path = os.path.join(tmp_dir_name, "encoder.onnx")
            Path(temp_onnx_model_path).parent.mkdir(parents=True, exist_ok=True)
            out_path = temp_onnx_model_path if use_external_data_format else onnx_model_path

            torch.onnx.export(
                self,
                args=(inputs["audio_features"]),
                f=out_path,
                export_params=True,
                input_names=input_names,
                output_names=output_names,
                dynamic_axes=dynamic_axes,
                opset_version=17,
                do_constant_folding=True,
                verbose=verbose,
            )

            model = onnx.load_model(out_path, load_external_data=use_external_data_format)
            model = self.fix_layernorm_weights(model, use_fp16_inputs)
            OnnxModel.save(
                model,
                onnx_model_path,
                save_as_external_data=use_external_data_format,
                all_tensors_to_one_file=True,
            )

        self.verify_onnx(onnx_model_path, provider, use_fp16_inputs)

    def verify_onnx(
        self,
        onnx_model_path: str,
        provider: str,
        use_fp16_inputs: bool,
    ):
        """Verify ONNX model outputs and PyTorch model outputs match

        Args:
            onnx_model_path (str): path to save ONNX model
            provider (str): execution provider for ONNX model
            use_fp16_inputs (bool, optional): use float16 inputs for the audio_features
        """
        # Shape of encoder's tensors:
        # Inputs:
        #    audio_features: (batch_size, num_mels, num_frames)
        # Outputs:
        #    encoder_hidden_states: (batch_size, num_frames // 2, hidden_size)
        inputs = get_sample_encoder_inputs(
            self.config,
            self.device,
            batch_size=2,
            use_fp16=use_fp16_inputs,
        )

        # Run PyTorch model
        pt_outputs = self.forward(inputs["audio_features"]).detach().cpu().numpy()

        # Run ONNX model
        sess = InferenceSession(onnx_model_path, providers=[provider])
        ort_outputs = sess.run(None, {"audio_features": inputs["audio_features"].detach().cpu().numpy()})[0]

        # Calculate output difference
        diff = np.abs(pt_outputs - ort_outputs)
        logger.warning("Comparing encoder_hidden_states...")
        logger.warning(f"Max diff: {np.max(diff)}")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\resolution\resolvelib\found_candidates.py ===
"""Utilities to lazily create and visit candidates found.

Creating and visiting a candidate is a *very* costly operation. It involves
fetching, extracting, potentially building modules from source, and verifying
distribution metadata. It is therefore crucial for performance to keep
everything here lazy all the way down, so we only touch candidates that we
absolutely need, and not "download the world" when we only need one version of
something.
"""

import logging
from collections.abc import Sequence
from typing import Any, Callable, Iterator, Optional, Set, Tuple

from pip._vendor.packaging.version import _BaseVersion

from pip._internal.exceptions import MetadataInvalid

from .base import Candidate

logger = logging.getLogger(__name__)

IndexCandidateInfo = Tuple[_BaseVersion, Callable[[], Optional[Candidate]]]


def _iter_built(infos: Iterator[IndexCandidateInfo]) -> Iterator[Candidate]:
    """Iterator for ``FoundCandidates``.

    This iterator is used when the package is not already installed. Candidates
    from index come later in their normal ordering.
    """
    versions_found: Set[_BaseVersion] = set()
    for version, func in infos:
        if version in versions_found:
            continue
        try:
            candidate = func()
        except MetadataInvalid as e:
            logger.warning(
                "Ignoring version %s of %s since it has invalid metadata:\n"
                "%s\n"
                "Please use pip<24.1 if you need to use this version.",
                version,
                e.ireq.name,
                e,
            )
            # Mark version as found to avoid trying other candidates with the same
            # version, since they most likely have invalid metadata as well.
            versions_found.add(version)
        else:
            if candidate is None:
                continue
            yield candidate
            versions_found.add(version)


def _iter_built_with_prepended(
    installed: Candidate, infos: Iterator[IndexCandidateInfo]
) -> Iterator[Candidate]:
    """Iterator for ``FoundCandidates``.

    This iterator is used when the resolver prefers the already-installed
    candidate and NOT to upgrade. The installed candidate is therefore
    always yielded first, and candidates from index come later in their
    normal ordering, except skipped when the version is already installed.
    """
    yield installed
    versions_found: Set[_BaseVersion] = {installed.version}
    for version, func in infos:
        if version in versions_found:
            continue
        candidate = func()
        if candidate is None:
            continue
        yield candidate
        versions_found.add(version)


def _iter_built_with_inserted(
    installed: Candidate, infos: Iterator[IndexCandidateInfo]
) -> Iterator[Candidate]:
    """Iterator for ``FoundCandidates``.

    This iterator is used when the resolver prefers to upgrade an
    already-installed package. Candidates from index are returned in their
    normal ordering, except replaced when the version is already installed.

    The implementation iterates through and yields other candidates, inserting
    the installed candidate exactly once before we start yielding older or
    equivalent candidates, or after all other candidates if they are all newer.
    """
    versions_found: Set[_BaseVersion] = set()
    for version, func in infos:
        if version in versions_found:
            continue
        # If the installed candidate is better, yield it first.
        if installed.version >= version:
            yield installed
            versions_found.add(installed.version)
        candidate = func()
        if candidate is None:
            continue
        yield candidate
        versions_found.add(version)

    # If the installed candidate is older than all other candidates.
    if installed.version not in versions_found:
        yield installed


class FoundCandidates(Sequence[Candidate]):
    """A lazy sequence to provide candidates to the resolver.

    The intended usage is to return this from `find_matches()` so the resolver
    can iterate through the sequence multiple times, but only access the index
    page when remote packages are actually needed. This improve performances
    when suitable candidates are already installed on disk.
    """

    def __init__(
        self,
        get_infos: Callable[[], Iterator[IndexCandidateInfo]],
        installed: Optional[Candidate],
        prefers_installed: bool,
        incompatible_ids: Set[int],
    ):
        self._get_infos = get_infos
        self._installed = installed
        self._prefers_installed = prefers_installed
        self._incompatible_ids = incompatible_ids
        self._bool: Optional[bool] = None

    def __getitem__(self, index: Any) -> Any:
        # Implemented to satisfy the ABC check. This is not needed by the
        # resolver, and should not be used by the provider either (for
        # performance reasons).
        raise NotImplementedError("don't do this")

    def __iter__(self) -> Iterator[Candidate]:
        infos = self._get_infos()
        if not self._installed:
            iterator = _iter_built(infos)
        elif self._prefers_installed:
            iterator = _iter_built_with_prepended(self._installed, infos)
        else:
            iterator = _iter_built_with_inserted(self._installed, infos)
        return (c for c in iterator if id(c) not in self._incompatible_ids)

    def __len__(self) -> int:
        # Implemented to satisfy the ABC check. This is not needed by the
        # resolver, and should not be used by the provider either (for
        # performance reasons).
        raise NotImplementedError("don't do this")

    def __bool__(self) -> bool:
        if self._bool is not None:
            return self._bool

        if self._prefers_installed and self._installed:
            self._bool = True
            return True

        self._bool = any(self)
        return self._bool

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pycparser\ast_transforms.py ===
#------------------------------------------------------------------------------
# pycparser: ast_transforms.py
#
# Some utilities used by the parser to create a friendlier AST.
#
# Eli Bendersky [https://eli.thegreenplace.net/]
# License: BSD
#------------------------------------------------------------------------------

from . import c_ast


def fix_switch_cases(switch_node):
    """ The 'case' statements in a 'switch' come out of parsing with one
        child node, so subsequent statements are just tucked to the parent
        Compound. Additionally, consecutive (fall-through) case statements
        come out messy. This is a peculiarity of the C grammar. The following:

            switch (myvar) {
                case 10:
                    k = 10;
                    p = k + 1;
                    return 10;
                case 20:
                case 30:
                    return 20;
                default:
                    break;
            }

        Creates this tree (pseudo-dump):

            Switch
                ID: myvar
                Compound:
                    Case 10:
                        k = 10
                    p = k + 1
                    return 10
                    Case 20:
                        Case 30:
                            return 20
                    Default:
                        break

        The goal of this transform is to fix this mess, turning it into the
        following:

            Switch
                ID: myvar
                Compound:
                    Case 10:
                        k = 10
                        p = k + 1
                        return 10
                    Case 20:
                    Case 30:
                        return 20
                    Default:
                        break

        A fixed AST node is returned. The argument may be modified.
    """
    assert isinstance(switch_node, c_ast.Switch)
    if not isinstance(switch_node.stmt, c_ast.Compound):
        return switch_node

    # The new Compound child for the Switch, which will collect children in the
    # correct order
    new_compound = c_ast.Compound([], switch_node.stmt.coord)

    # The last Case/Default node
    last_case = None

    # Goes over the children of the Compound below the Switch, adding them
    # either directly below new_compound or below the last Case as appropriate
    # (for `switch(cond) {}`, block_items would have been None)
    for child in (switch_node.stmt.block_items or []):
        if isinstance(child, (c_ast.Case, c_ast.Default)):
            # If it's a Case/Default:
            # 1. Add it to the Compound and mark as "last case"
            # 2. If its immediate child is also a Case or Default, promote it
            #    to a sibling.
            new_compound.block_items.append(child)
            _extract_nested_case(child, new_compound.block_items)
            last_case = new_compound.block_items[-1]
        else:
            # Other statements are added as children to the last case, if it
            # exists.
            if last_case is None:
                new_compound.block_items.append(child)
            else:
                last_case.stmts.append(child)

    switch_node.stmt = new_compound
    return switch_node


def _extract_nested_case(case_node, stmts_list):
    """ Recursively extract consecutive Case statements that are made nested
        by the parser and add them to the stmts_list.
    """
    if isinstance(case_node.stmts[0], (c_ast.Case, c_ast.Default)):
        stmts_list.append(case_node.stmts.pop())
        _extract_nested_case(stmts_list[-1], stmts_list)


def fix_atomic_specifiers(decl):
    """ Atomic specifiers like _Atomic(type) are unusually structured,
        conferring a qualifier upon the contained type.

        This function fixes a decl with atomic specifiers to have a sane AST
        structure, by removing spurious Typename->TypeDecl pairs and attaching
        the _Atomic qualifier in the right place.
    """
    # There can be multiple levels of _Atomic in a decl; fix them until a
    # fixed point is reached.
    while True:
        decl, found = _fix_atomic_specifiers_once(decl)
        if not found:
            break

    # Make sure to add an _Atomic qual on the topmost decl if needed. Also
    # restore the declname on the innermost TypeDecl (it gets placed in the
    # wrong place during construction).
    typ = decl
    while not isinstance(typ, c_ast.TypeDecl):
        try:
            typ = typ.type
        except AttributeError:
            return decl
    if '_Atomic' in typ.quals and '_Atomic' not in decl.quals:
        decl.quals.append('_Atomic')
    if typ.declname is None:
        typ.declname = decl.name

    return decl


def _fix_atomic_specifiers_once(decl):
    """ Performs one 'fix' round of atomic specifiers.
        Returns (modified_decl, found) where found is True iff a fix was made.
    """
    parent = decl
    grandparent = None
    node = decl.type
    while node is not None:
        if isinstance(node, c_ast.Typename) and '_Atomic' in node.quals:
            break
        try:
            grandparent = parent
            parent = node
            node = node.type
        except AttributeError:
            # If we've reached a node without a `type` field, it means we won't
            # find what we're looking for at this point; give up the search
            # and return the original decl unmodified.
            return decl, False

    assert isinstance(parent, c_ast.TypeDecl)
    grandparent.type = node.type
    if '_Atomic' not in node.type.quals:
        node.type.quals.append('_Atomic')
    return decl, True

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\orm\sync.py ===
# orm/sync.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: allow-untyped-defs, allow-untyped-calls


"""private module containing functions used for copying data
between instances based on join conditions.

"""

from __future__ import annotations

from . import exc
from . import util as orm_util
from .base import PassiveFlag


def populate(
    source,
    source_mapper,
    dest,
    dest_mapper,
    synchronize_pairs,
    uowcommit,
    flag_cascaded_pks,
):
    source_dict = source.dict
    dest_dict = dest.dict

    for l, r in synchronize_pairs:
        try:
            # inline of source_mapper._get_state_attr_by_column
            prop = source_mapper._columntoproperty[l]
            value = source.manager[prop.key].impl.get(
                source, source_dict, PassiveFlag.PASSIVE_OFF
            )
        except exc.UnmappedColumnError as err:
            _raise_col_to_prop(False, source_mapper, l, dest_mapper, r, err)

        try:
            # inline of dest_mapper._set_state_attr_by_column
            prop = dest_mapper._columntoproperty[r]
            dest.manager[prop.key].impl.set(dest, dest_dict, value, None)
        except exc.UnmappedColumnError as err:
            _raise_col_to_prop(True, source_mapper, l, dest_mapper, r, err)

        # technically the "r.primary_key" check isn't
        # needed here, but we check for this condition to limit
        # how often this logic is invoked for memory/performance
        # reasons, since we only need this info for a primary key
        # destination.
        if (
            flag_cascaded_pks
            and l.primary_key
            and r.primary_key
            and r.references(l)
        ):
            uowcommit.attributes[("pk_cascaded", dest, r)] = True


def bulk_populate_inherit_keys(source_dict, source_mapper, synchronize_pairs):
    # a simplified version of populate() used by bulk insert mode
    for l, r in synchronize_pairs:
        try:
            prop = source_mapper._columntoproperty[l]
            value = source_dict[prop.key]
        except exc.UnmappedColumnError as err:
            _raise_col_to_prop(False, source_mapper, l, source_mapper, r, err)

        try:
            prop = source_mapper._columntoproperty[r]
            source_dict[prop.key] = value
        except exc.UnmappedColumnError as err:
            _raise_col_to_prop(True, source_mapper, l, source_mapper, r, err)


def clear(dest, dest_mapper, synchronize_pairs):
    for l, r in synchronize_pairs:
        if (
            r.primary_key
            and dest_mapper._get_state_attr_by_column(dest, dest.dict, r)
            not in orm_util._none_set
        ):
            raise AssertionError(
                f"Dependency rule on column '{l}' "
                "tried to blank-out primary key "
                f"column '{r}' on instance '{orm_util.state_str(dest)}'"
            )
        try:
            dest_mapper._set_state_attr_by_column(dest, dest.dict, r, None)
        except exc.UnmappedColumnError as err:
            _raise_col_to_prop(True, None, l, dest_mapper, r, err)


def update(source, source_mapper, dest, old_prefix, synchronize_pairs):
    for l, r in synchronize_pairs:
        try:
            oldvalue = source_mapper._get_committed_attr_by_column(
                source.obj(), l
            )
            value = source_mapper._get_state_attr_by_column(
                source, source.dict, l, passive=PassiveFlag.PASSIVE_OFF
            )
        except exc.UnmappedColumnError as err:
            _raise_col_to_prop(False, source_mapper, l, None, r, err)
        dest[r.key] = value
        dest[old_prefix + r.key] = oldvalue


def populate_dict(source, source_mapper, dict_, synchronize_pairs):
    for l, r in synchronize_pairs:
        try:
            value = source_mapper._get_state_attr_by_column(
                source, source.dict, l, passive=PassiveFlag.PASSIVE_OFF
            )
        except exc.UnmappedColumnError as err:
            _raise_col_to_prop(False, source_mapper, l, None, r, err)

        dict_[r.key] = value


def source_modified(uowcommit, source, source_mapper, synchronize_pairs):
    """return true if the source object has changes from an old to a
    new value on the given synchronize pairs

    """
    for l, r in synchronize_pairs:
        try:
            prop = source_mapper._columntoproperty[l]
        except exc.UnmappedColumnError as err:
            _raise_col_to_prop(False, source_mapper, l, None, r, err)
        history = uowcommit.get_attribute_history(
            source, prop.key, PassiveFlag.PASSIVE_NO_INITIALIZE
        )
        if bool(history.deleted):
            return True
    else:
        return False


def _raise_col_to_prop(
    isdest, source_mapper, source_column, dest_mapper, dest_column, err
):
    if isdest:
        raise exc.UnmappedColumnError(
            "Can't execute sync rule for "
            "destination column '%s'; mapper '%s' does not map "
            "this column.  Try using an explicit `foreign_keys` "
            "collection which does not include this column (or use "
            "a viewonly=True relation)." % (dest_column, dest_mapper)
        ) from err
    else:
        raise exc.UnmappedColumnError(
            "Can't execute sync rule for "
            "source column '%s'; mapper '%s' does not map this "
            "column.  Try using an explicit `foreign_keys` "
            "collection which does not include destination column "
            "'%s' (or use a viewonly=True relation)."
            % (source_column, source_mapper, dest_column)
        ) from err

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\assumptions\wrapper.py ===
"""
Functions and wrapper object to call assumption property and predicate
query with same syntax.

In SymPy, there are two assumption systems. Old assumption system is
defined in sympy/core/assumptions, and it can be accessed by attribute
such as ``x.is_even``. New assumption system is defined in
sympy/assumptions, and it can be accessed by predicates such as
``Q.even(x)``.

Old assumption is fast, while new assumptions can freely take local facts.
In general, old assumption is used in evaluation method and new assumption
is used in refinement method.

In most cases, both evaluation and refinement follow the same process, and
the only difference is which assumption system is used. This module provides
``is_[...]()`` functions and ``AssumptionsWrapper()`` class which allows
using two systems with same syntax so that parallel code implementation can be
avoided.

Examples
========

For multiple use, use ``AssumptionsWrapper()``.

>>> from sympy import Q, Symbol
>>> from sympy.assumptions.wrapper import AssumptionsWrapper
>>> x = Symbol('x')
>>> _x = AssumptionsWrapper(x, Q.even(x))
>>> _x.is_integer
True
>>> _x.is_odd
False

For single use, use ``is_[...]()`` functions.

>>> from sympy.assumptions.wrapper import is_infinite
>>> a = Symbol('a')
>>> print(is_infinite(a))
None
>>> is_infinite(a, Q.finite(a))
False

"""

from sympy.assumptions import ask, Q
from sympy.core.basic import Basic
from sympy.core.sympify import _sympify


def make_eval_method(fact):
    def getit(self):
        pred = getattr(Q, fact)
        ret = ask(pred(self.expr), self.assumptions)
        return ret
    return getit


# we subclass Basic to use the fact deduction and caching
class AssumptionsWrapper(Basic):
    """
    Wrapper over ``Basic`` instances to call predicate query by
    ``.is_[...]`` property

    Parameters
    ==========

    expr : Basic

    assumptions : Boolean, optional

    Examples
    ========

    >>> from sympy import Q, Symbol
    >>> from sympy.assumptions.wrapper import AssumptionsWrapper
    >>> x = Symbol('x', even=True)
    >>> AssumptionsWrapper(x).is_integer
    True
    >>> y = Symbol('y')
    >>> AssumptionsWrapper(y, Q.even(y)).is_integer
    True

    With ``AssumptionsWrapper``, both evaluation and refinement can be supported
    by single implementation.

    >>> from sympy import Function
    >>> class MyAbs(Function):
    ...     @classmethod
    ...     def eval(cls, x, assumptions=True):
    ...         _x = AssumptionsWrapper(x, assumptions)
    ...         if _x.is_nonnegative:
    ...             return x
    ...         if _x.is_negative:
    ...             return -x
    ...     def _eval_refine(self, assumptions):
    ...         return MyAbs.eval(self.args[0], assumptions)
    >>> MyAbs(x)
    MyAbs(x)
    >>> MyAbs(x).refine(Q.positive(x))
    x
    >>> MyAbs(Symbol('y', negative=True))
    -y

    """
    def __new__(cls, expr, assumptions=None):
        if assumptions is None:
            return expr
        obj = super().__new__(cls, expr, _sympify(assumptions))
        obj.expr = expr
        obj.assumptions = assumptions
        return obj

    _eval_is_algebraic = make_eval_method("algebraic")
    _eval_is_antihermitian = make_eval_method("antihermitian")
    _eval_is_commutative = make_eval_method("commutative")
    _eval_is_complex = make_eval_method("complex")
    _eval_is_composite = make_eval_method("composite")
    _eval_is_even = make_eval_method("even")
    _eval_is_extended_negative = make_eval_method("extended_negative")
    _eval_is_extended_nonnegative = make_eval_method("extended_nonnegative")
    _eval_is_extended_nonpositive = make_eval_method("extended_nonpositive")
    _eval_is_extended_nonzero = make_eval_method("extended_nonzero")
    _eval_is_extended_positive = make_eval_method("extended_positive")
    _eval_is_extended_real = make_eval_method("extended_real")
    _eval_is_finite = make_eval_method("finite")
    _eval_is_hermitian = make_eval_method("hermitian")
    _eval_is_imaginary = make_eval_method("imaginary")
    _eval_is_infinite = make_eval_method("infinite")
    _eval_is_integer = make_eval_method("integer")
    _eval_is_irrational = make_eval_method("irrational")
    _eval_is_negative = make_eval_method("negative")
    _eval_is_noninteger = make_eval_method("noninteger")
    _eval_is_nonnegative = make_eval_method("nonnegative")
    _eval_is_nonpositive = make_eval_method("nonpositive")
    _eval_is_nonzero = make_eval_method("nonzero")
    _eval_is_odd = make_eval_method("odd")
    _eval_is_polar = make_eval_method("polar")
    _eval_is_positive = make_eval_method("positive")
    _eval_is_prime = make_eval_method("prime")
    _eval_is_rational = make_eval_method("rational")
    _eval_is_real = make_eval_method("real")
    _eval_is_transcendental = make_eval_method("transcendental")
    _eval_is_zero = make_eval_method("zero")


# one shot functions which are faster than AssumptionsWrapper

def is_infinite(obj, assumptions=None):
    if assumptions is None:
        return obj.is_infinite
    return ask(Q.infinite(obj), assumptions)


def is_extended_real(obj, assumptions=None):
    if assumptions is None:
        return obj.is_extended_real
    return ask(Q.extended_real(obj), assumptions)


def is_extended_nonnegative(obj, assumptions=None):
    if assumptions is None:
        return obj.is_extended_nonnegative
    return ask(Q.extended_nonnegative(obj), assumptions)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\assumptions\handlers\common.py ===
"""
This module defines base class for handlers and some core handlers:
``Q.commutative`` and ``Q.is_true``.
"""

from sympy.assumptions import Q, ask, AppliedPredicate
from sympy.core import Basic, Symbol
from sympy.core.logic import _fuzzy_group, fuzzy_and, fuzzy_or
from sympy.core.numbers import NaN, Number
from sympy.logic.boolalg import (And, BooleanTrue, BooleanFalse, conjuncts,
    Equivalent, Implies, Not, Or)
from sympy.utilities.exceptions import sympy_deprecation_warning

from ..predicates.common import CommutativePredicate, IsTruePredicate


class AskHandler:
    """Base class that all Ask Handlers must inherit."""
    def __new__(cls, *args, **kwargs):
        sympy_deprecation_warning(
            """
            The AskHandler system is deprecated. The AskHandler class should
            be replaced with the multipledispatch handler of Predicate
            """,
            deprecated_since_version="1.8",
            active_deprecations_target='deprecated-askhandler',
        )
        return super().__new__(cls, *args, **kwargs)


class CommonHandler(AskHandler):
    # Deprecated
    """Defines some useful methods common to most Handlers. """

    @staticmethod
    def AlwaysTrue(expr, assumptions):
        return True

    @staticmethod
    def AlwaysFalse(expr, assumptions):
        return False

    @staticmethod
    def AlwaysNone(expr, assumptions):
        return None

    NaN = AlwaysFalse


# CommutativePredicate

@CommutativePredicate.register(Symbol)
def _(expr, assumptions):
    """Objects are expected to be commutative unless otherwise stated"""
    assumps = conjuncts(assumptions)
    if expr.is_commutative is not None:
        return expr.is_commutative and not ~Q.commutative(expr) in assumps
    if Q.commutative(expr) in assumps:
        return True
    elif ~Q.commutative(expr) in assumps:
        return False
    return True

@CommutativePredicate.register(Basic)
def _(expr, assumptions):
    for arg in expr.args:
        if not ask(Q.commutative(arg), assumptions):
            return False
    return True

@CommutativePredicate.register(Number)
def _(expr, assumptions):
    return True

@CommutativePredicate.register(NaN)
def _(expr, assumptions):
    return True


# IsTruePredicate

@IsTruePredicate.register(bool)
def _(expr, assumptions):
    return expr

@IsTruePredicate.register(BooleanTrue)
def _(expr, assumptions):
    return True

@IsTruePredicate.register(BooleanFalse)
def _(expr, assumptions):
    return False

@IsTruePredicate.register(AppliedPredicate)
def _(expr, assumptions):
    return ask(expr, assumptions)

@IsTruePredicate.register(Not)
def _(expr, assumptions):
    arg = expr.args[0]
    if arg.is_Symbol:
        # symbol used as abstract boolean object
        return None
    value = ask(arg, assumptions=assumptions)
    if value in (True, False):
        return not value
    else:
        return None

@IsTruePredicate.register(Or)
def _(expr, assumptions):
    result = False
    for arg in expr.args:
        p = ask(arg, assumptions=assumptions)
        if p is True:
            return True
        if p is None:
            result = None
    return result

@IsTruePredicate.register(And)
def _(expr, assumptions):
    result = True
    for arg in expr.args:
        p = ask(arg, assumptions=assumptions)
        if p is False:
            return False
        if p is None:
            result = None
    return result

@IsTruePredicate.register(Implies)
def _(expr, assumptions):
    p, q = expr.args
    return ask(~p | q, assumptions=assumptions)

@IsTruePredicate.register(Equivalent)
def _(expr, assumptions):
    p, q = expr.args
    pt = ask(p, assumptions=assumptions)
    if pt is None:
        return None
    qt = ask(q, assumptions=assumptions)
    if qt is None:
        return None
    return pt == qt


#### Helper methods
def test_closed_group(expr, assumptions, key):
    """
    Test for membership in a group with respect
    to the current operation.
    """
    return _fuzzy_group(
        (ask(key(a), assumptions) for a in expr.args), quick_exit=True)

def ask_all(*queries, assumptions):
    return fuzzy_and(
        (ask(query, assumptions) for query in queries))

def ask_any(*queries, assumptions):
    return fuzzy_or(
        (ask(query, assumptions) for query in queries))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\liealgebras\type_a.py ===
from sympy.liealgebras.cartan_type import Standard_Cartan
from sympy.core.backend import eye


class TypeA(Standard_Cartan):
    """
    This class contains the information about
    the A series of simple Lie algebras.
    ====
    """

    def __new__(cls, n):
        if n < 1:
            raise ValueError("n cannot be less than 1")
        return Standard_Cartan.__new__(cls, "A", n)


    def dimension(self):
        """Dimension of the vector space V underlying the Lie algebra

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType("A4")
        >>> c.dimension()
        5
        """
        return self.n+1


    def basic_root(self, i, j):
        """
        This is a method just to generate roots
        with a 1 iin the ith position and a -1
        in the jth position.

        """

        n = self.n
        root = [0]*(n+1)
        root[i] = 1
        root[j] = -1
        return root

    def simple_root(self, i):
        """
        Every lie algebra has a unique root system.
        Given a root system Q, there is a subset of the
        roots such that an element of Q is called a
        simple root if it cannot be written as the sum
        of two elements in Q.  If we let D denote the
        set of simple roots, then it is clear that every
        element of Q can be written as a linear combination
        of elements of D with all coefficients non-negative.

        In A_n the ith simple root is the root which has a 1
        in the ith position, a -1 in the (i+1)th position,
        and zeroes elsewhere.

        This method returns the ith simple root for the A series.

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType("A4")
        >>> c.simple_root(1)
        [1, -1, 0, 0, 0]

        """

        return self.basic_root(i-1, i)

    def positive_roots(self):
        """
        This method generates all the positive roots of
        A_n.  This is half of all of the roots of A_n;
        by multiplying all the positive roots by -1 we
        get the negative roots.

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType("A3")
        >>> c.positive_roots()
        {1: [1, -1, 0, 0], 2: [1, 0, -1, 0], 3: [1, 0, 0, -1], 4: [0, 1, -1, 0],
                5: [0, 1, 0, -1], 6: [0, 0, 1, -1]}
        """

        n = self.n
        posroots = {}
        k = 0
        for i in range(0, n):
            for j in range(i+1, n+1):
                k += 1
                posroots[k] = self.basic_root(i, j)
        return posroots

    def highest_root(self):
        """
        Returns the highest weight root for A_n
        """

        return self.basic_root(0, self.n)

    def roots(self):
        """
        Returns the total number of roots for A_n
        """
        n = self.n
        return n*(n+1)

    def cartan_matrix(self):
        """
        Returns the Cartan matrix for A_n.
        The Cartan matrix matrix for a Lie algebra is
        generated by assigning an ordering to the simple
        roots, (alpha[1], ...., alpha[l]).  Then the ijth
        entry of the Cartan matrix is (<alpha[i],alpha[j]>).

        Examples
        ========

        >>> from sympy.liealgebras.cartan_type import CartanType
        >>> c = CartanType('A4')
        >>> c.cartan_matrix()
        Matrix([
        [ 2, -1,  0,  0],
        [-1,  2, -1,  0],
        [ 0, -1,  2, -1],
        [ 0,  0, -1,  2]])

        """

        n = self.n
        m = 2 * eye(n)
        for i in range(1, n - 1):
            m[i, i+1] = -1
            m[i, i-1] = -1
        m[0,1] = -1
        m[n-1, n-2] = -1
        return m

    def basis(self):
        """
        Returns the number of independent generators of A_n
        """
        n = self.n
        return n**2 - 1

    def lie_algebra(self):
        """
        Returns the Lie algebra associated with A_n
        """
        n = self.n
        return "su(" + str(n + 1) + ")"

    def dynkin_diagram(self):
        n = self.n
        diag = "---".join("0" for i in range(1, n+1)) + "\n"
        diag += "   ".join(str(i) for i in range(1, n+1))
        return diag

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\webdriver_manager\core\driver_cache.py ===
import datetime
import json
import os

from webdriver_manager.core.config import wdm_local, get_xdist_worker_id
from webdriver_manager.core.constants import (
    DEFAULT_PROJECT_ROOT_CACHE_PATH,
    DEFAULT_USER_HOME_CACHE_PATH, ROOT_FOLDER_NAME,
)
from webdriver_manager.core.driver import Driver
from webdriver_manager.core.file_manager import FileManager, File
from webdriver_manager.core.logger import log
from webdriver_manager.core.os_manager import OperationSystemManager
from webdriver_manager.core.utils import get_date_diff


class DriverCacheManager(object):
    def __init__(self, root_dir=None, valid_range=1, file_manager=None):
        self._root_dir = DEFAULT_USER_HOME_CACHE_PATH
        is_wdm_local = wdm_local()
        xdist_worker_id = get_xdist_worker_id()
        if xdist_worker_id:
            log(f"xdist worker is: {xdist_worker_id}")
            self._root_dir = os.path.join(self._root_dir, xdist_worker_id)

        if root_dir:
            self._root_dir = os.path.join(root_dir, ROOT_FOLDER_NAME, xdist_worker_id)

        if is_wdm_local:
            self._root_dir = os.path.join(DEFAULT_PROJECT_ROOT_CACHE_PATH, xdist_worker_id)

        self._drivers_root = "drivers"
        self._drivers_json_path = os.path.join(self._root_dir, "drivers.json")
        self._date_format = "%d/%m/%Y"
        self._drivers_directory = os.path.join(self._root_dir, self._drivers_root)
        self._cache_valid_days_range = valid_range
        self._cache_key_driver_version = None
        self._metadata_key = None
        self._driver_binary_path = None
        self._file_manager = file_manager
        self._os_system_manager = OperationSystemManager()
        if not self._file_manager:
            self._file_manager = FileManager(self._os_system_manager)

    def save_archive_file(self, file: File, path):
        return self._file_manager.save_archive_file(file, path)

    def unpack_archive(self, archive, path):
        return self._file_manager.unpack_archive(archive, path)

    def save_file_to_cache(self, driver: Driver, file: File):
        path = self.__get_path(driver)
        archive = self.save_archive_file(file, path)
        files = self.unpack_archive(archive, path)
        binary = self.__get_binary(files, driver.get_name())
        binary_path = os.path.join(path, binary)
        self.__save_metadata(driver, binary_path)
        log(f"Driver has been saved in cache [{path}]")
        return binary_path

    def __get_binary(self, files, driver_name):
        if not files:
            raise Exception(f"Can't find binary for {driver_name} among {files}")

        if len(files) == 1:
            return files[0]

        for f in files:
            if 'LICENSE' in f:
                continue
            if 'THIRD_PARTY' in f:
                continue
            if driver_name in f:
                return f

        raise Exception(f"Can't find binary for {driver_name} among {files}")

    def __save_metadata(self, driver: Driver, binary_path, date=None):
        if date is None:
            date = datetime.date.today()

        metadata = self.load_metadata_content()
        key = self.__get_metadata_key(driver)
        data = {
            key: {
                "timestamp": date.strftime(self._date_format),
                "binary_path": binary_path,
            }
        }

        metadata.update(data)
        with open(self._drivers_json_path, "w+") as outfile:
            json.dump(metadata, outfile, indent=4)

    def get_os_type(self):
        return self._os_system_manager.get_os_type()

    def find_driver(self, driver: Driver):
        """Find driver by '{os_type}_{driver_name}_{driver_version}_{browser_version}'."""
        os_type = self.get_os_type()
        driver_name = driver.get_name()
        browser_type = driver.get_browser_type()
        browser_version = self._os_system_manager.get_browser_version_from_os(browser_type)
        if not browser_version:
            return None

        driver_version = self.get_cache_key_driver_version(driver)
        metadata = self.load_metadata_content()

        key = self.__get_metadata_key(driver)
        if key not in metadata:
            log(f'There is no [{os_type}] {driver_name} "{driver_version}" for browser {browser_type} '
                f'"{browser_version}" in cache')
            return None

        driver_info = metadata[key]
        path = driver_info["binary_path"]
        if not os.path.exists(path):
            return None

        if not self.__is_valid(driver_info):
            return None

        path = driver_info["binary_path"]
        log(f"Driver [{path}] found in cache")
        return path

    def __is_valid(self, driver_info):
        dates_diff = get_date_diff(
            driver_info["timestamp"], datetime.date.today(), self._date_format
        )
        return dates_diff < self._cache_valid_days_range

    def load_metadata_content(self):
        if os.path.exists(self._drivers_json_path):
            with open(self._drivers_json_path, "r") as outfile:
                return json.load(outfile)
        return {}

    def __get_metadata_key(self, driver: Driver):
        if self._metadata_key:
            return self._metadata_key

        driver_version = self.get_cache_key_driver_version(driver)
        browser_version = driver.get_browser_version_from_os()
        browser_version = browser_version if browser_version else ""
        self._metadata_key = f"{self.get_os_type()}_{driver.get_name()}_{driver_version}" \
                             f"_for_{browser_version}"
        return self._metadata_key

    def get_cache_key_driver_version(self, driver: Driver):
        if self._cache_key_driver_version:
            return self._cache_key_driver_version
        return driver.get_driver_version_to_download()

    def __get_path(self, driver: Driver):
        if self._driver_binary_path is None:
            self._driver_binary_path = os.path.join(
                self._drivers_directory,
                driver.get_name(),
                self.get_os_type(),
                driver.get_driver_version_to_download(),
            )
        return self._driver_binary_path

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\parsing\tests\test_c_parser.py ===
from sympy.parsing.sym_expr import SymPyExpression
from sympy.testing.pytest import raises, XFAIL
from sympy.external import import_module

cin = import_module('clang.cindex', import_kwargs = {'fromlist': ['cindex']})

if cin:
    from sympy.codegen.ast import (Variable, String, Return,
        FunctionDefinition, Integer, Float, Declaration, CodeBlock,
        FunctionPrototype, FunctionCall, NoneToken, Assignment, Type,
        IntBaseType, SignedIntType, UnsignedIntType, FloatType,
        AddAugmentedAssignment, SubAugmentedAssignment,
        MulAugmentedAssignment, DivAugmentedAssignment,
        ModAugmentedAssignment, While)
    from sympy.codegen.cnodes import (PreDecrement, PostDecrement,
        PreIncrement, PostIncrement)
    from sympy.core import (Add, Mul, Mod, Pow, Rational,
        StrictLessThan, LessThan, StrictGreaterThan, GreaterThan,
        Equality, Unequality)
    from sympy.logic.boolalg import And, Not, Or
    from sympy.core.symbol import Symbol
    from sympy.logic.boolalg import (false, true)
    import os

    def test_variable():
        c_src1 = (
            'int a;' + '\n' +
            'int b;' + '\n'
        )
        c_src2 = (
            'float a;' + '\n'
            + 'float b;' + '\n'
        )
        c_src3 = (
            'int a;' + '\n' +
            'float b;' + '\n' +
            'int c;'
        )
        c_src4 = (
            'int x = 1, y = 6.78;' + '\n' +
            'float p = 2, q = 9.67;'
        )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()
        res3 = SymPyExpression(c_src3, 'c').return_expr()
        res4 = SymPyExpression(c_src4, 'c').return_expr()

        assert res1[0] == Declaration(
            Variable(
                Symbol('a'),
                type=IntBaseType(String('intc'))
            )
        )

        assert res1[1] == Declaration(
            Variable(
                Symbol('b'),
                type=IntBaseType(String('intc'))
            )
        )

        assert res2[0] == Declaration(
            Variable(
                Symbol('a'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    )
            )
        )
        assert res2[1] == Declaration(
            Variable(
                Symbol('b'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    )
            )
        )

        assert res3[0] == Declaration(
            Variable(
                Symbol('a'),
                type=IntBaseType(String('intc'))
            )
        )

        assert res3[1] == Declaration(
            Variable(
                Symbol('b'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    )
            )
        )

        assert res3[2] == Declaration(
            Variable(
                Symbol('c'),
                type=IntBaseType(String('intc'))
            )
        )

        assert res4[0] == Declaration(
            Variable(
                Symbol('x'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
            )
        )

        assert res4[1] == Declaration(
            Variable(
                Symbol('y'),
                type=IntBaseType(String('intc')),
                value=Integer(6)
            )
        )

        assert res4[2] == Declaration(
            Variable(
                Symbol('p'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('2.0', precision=53)
            )
        )

        assert res4[3] == Declaration(
            Variable(
                Symbol('q'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('9.67', precision=53)
            )
        )


    def test_int():
        c_src1 = 'int a = 1;'
        c_src2 = (
            'int a = 1;' + '\n' +
            'int b = 2;' + '\n'
        )
        c_src3 = 'int a = 2.345, b = 5.67;'
        c_src4 = 'int p = 6, q = 23.45;'
        c_src5 = "int x = '0', y = 'a';"
        c_src6 = "int r = true, s = false;"

        # cin.TypeKind.UCHAR
        c_src_type1 = (
            "signed char a = 1, b = 5.1;"
            )

        # cin.TypeKind.SHORT
        c_src_type2 = (
            "short a = 1, b = 5.1;"
            "signed short c = 1, d = 5.1;"
            "short int e = 1, f = 5.1;"
            "signed short int g = 1, h = 5.1;"
            )

        # cin.TypeKind.INT
        c_src_type3 = (
            "signed int a = 1, b = 5.1;"
            "int c = 1, d = 5.1;"
            )

        # cin.TypeKind.LONG
        c_src_type4 = (
            "long a = 1, b = 5.1;"
            "long int c = 1, d = 5.1;"
            )

        # cin.TypeKind.UCHAR
        c_src_type5 = "unsigned char a = 1, b = 5.1;"

        # cin.TypeKind.USHORT
        c_src_type6 = (
            "unsigned short a = 1, b = 5.1;"
            "unsigned short int c = 1, d = 5.1;"
            )

        # cin.TypeKind.UINT
        c_src_type7 = "unsigned int a = 1, b = 5.1;"

        # cin.TypeKind.ULONG
        c_src_type8 = (
            "unsigned long a = 1, b = 5.1;"
            "unsigned long int c = 1, d = 5.1;"
            )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()
        res3 = SymPyExpression(c_src3, 'c').return_expr()
        res4 = SymPyExpression(c_src4, 'c').return_expr()
        res5 = SymPyExpression(c_src5, 'c').return_expr()
        res6 = SymPyExpression(c_src6, 'c').return_expr()

        res_type1 = SymPyExpression(c_src_type1, 'c').return_expr()
        res_type2 = SymPyExpression(c_src_type2, 'c').return_expr()
        res_type3 = SymPyExpression(c_src_type3, 'c').return_expr()
        res_type4 = SymPyExpression(c_src_type4, 'c').return_expr()
        res_type5 = SymPyExpression(c_src_type5, 'c').return_expr()
        res_type6 = SymPyExpression(c_src_type6, 'c').return_expr()
        res_type7 = SymPyExpression(c_src_type7, 'c').return_expr()
        res_type8 = SymPyExpression(c_src_type8, 'c').return_expr()

        assert res1[0] == Declaration(
            Variable(
                Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
            )
        )

        assert res2[0] == Declaration(
            Variable(
                Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
            )
        )

        assert res2[1] == Declaration(
            Variable(
                Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(2)
            )
        )

        assert res3[0] == Declaration(
            Variable(
                Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(2)
            )
        )

        assert res3[1] == Declaration(
            Variable(
                Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(5)
            )
        )

        assert res4[0] == Declaration(
            Variable(
                Symbol('p'),
                type=IntBaseType(String('intc')),
                value=Integer(6)
            )
        )

        assert res4[1] == Declaration(
            Variable(
                Symbol('q'),
                type=IntBaseType(String('intc')),
                value=Integer(23)
            )
        )

        assert res5[0] == Declaration(
            Variable(
                Symbol('x'),
                type=IntBaseType(String('intc')),
                value=Integer(48)
            )
        )

        assert res5[1] == Declaration(
            Variable(
                Symbol('y'),
                type=IntBaseType(String('intc')),
                value=Integer(97)
            )
        )

        assert res6[0] == Declaration(
            Variable(
                Symbol('r'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
            )
        )

        assert res6[1] == Declaration(
            Variable(
                Symbol('s'),
                type=IntBaseType(String('intc')),
                value=Integer(0)
            )
        )

        assert res_type1[0] == Declaration(
            Variable(
                Symbol('a'),
                type=SignedIntType(
                    String('int8'),
                    nbits=Integer(8)
                    ),
                value=Integer(1)
                )
            )

        assert res_type1[1] == Declaration(
            Variable(
                Symbol('b'),
                type=SignedIntType(
                    String('int8'),
                    nbits=Integer(8)
                    ),
                value=Integer(5)
                )
            )

        assert res_type2[0] == Declaration(
            Variable(
                Symbol('a'),
                type=SignedIntType(
                    String('int16'),
                    nbits=Integer(16)
                    ),
                value=Integer(1)
                )
            )

        assert res_type2[1] == Declaration(
            Variable(
                Symbol('b'),
                type=SignedIntType(
                    String('int16'),
                    nbits=Integer(16)
                    ),
                value=Integer(5)
                )
            )

        assert res_type2[2] == Declaration(
            Variable(Symbol('c'),
                type=SignedIntType(
                    String('int16'),
                    nbits=Integer(16)
                    ),
                value=Integer(1)
                )
            )

        assert res_type2[3] == Declaration(
            Variable(
                Symbol('d'),
                type=SignedIntType(
                    String('int16'),
                    nbits=Integer(16)
                    ),
                value=Integer(5)
                )
            )

        assert res_type2[4] == Declaration(
            Variable(
                Symbol('e'),
                type=SignedIntType(
                    String('int16'),
                    nbits=Integer(16)
                    ),
                value=Integer(1)
                )
            )

        assert res_type2[5] == Declaration(
            Variable(
                Symbol('f'),
                type=SignedIntType(
                    String('int16'),
                    nbits=Integer(16)
                    ),
                value=Integer(5)
                )
            )

        assert res_type2[6] == Declaration(
            Variable(
                Symbol('g'),
                type=SignedIntType(
                    String('int16'),
                    nbits=Integer(16)
                    ),
                value=Integer(1)
                )
            )

        assert res_type2[7] == Declaration(
            Variable(
                Symbol('h'),
                type=SignedIntType(
                    String('int16'),
                    nbits=Integer(16)
                    ),
                value=Integer(5)
                )
            )

        assert res_type3[0] == Declaration(
            Variable(
                Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res_type3[1] == Declaration(
            Variable(
                Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(5)
                )
            )

        assert res_type3[2] == Declaration(
            Variable(
                Symbol('c'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res_type3[3] == Declaration(
            Variable(
                Symbol('d'),
                type=IntBaseType(String('intc')),
                value=Integer(5)
                )
            )

        assert res_type4[0] == Declaration(
            Variable(
                Symbol('a'),
                type=SignedIntType(
                    String('int64'),
                    nbits=Integer(64)
                    ),
                value=Integer(1)
                )
            )

        assert res_type4[1] == Declaration(
            Variable(
                Symbol('b'),
                type=SignedIntType(
                    String('int64'),
                    nbits=Integer(64)
                    ),
                value=Integer(5)
                )
            )

        assert res_type4[2] == Declaration(
            Variable(
                Symbol('c'),
                type=SignedIntType(
                    String('int64'),
                    nbits=Integer(64)
                    ),
                value=Integer(1)
                )
            )

        assert res_type4[3] == Declaration(
            Variable(
                Symbol('d'),
                type=SignedIntType(
                    String('int64'),
                    nbits=Integer(64)
                    ),
                value=Integer(5)
                )
            )

        assert res_type5[0] == Declaration(
            Variable(
                Symbol('a'),
                type=UnsignedIntType(
                    String('uint8'),
                    nbits=Integer(8)
                    ),
                value=Integer(1)
                )
            )

        assert res_type5[1] == Declaration(
            Variable(
                Symbol('b'),
                type=UnsignedIntType(
                    String('uint8'),
                    nbits=Integer(8)
                    ),
                value=Integer(5)
                )
            )

        assert res_type6[0] == Declaration(
            Variable(
                Symbol('a'),
                type=UnsignedIntType(
                    String('uint16'),
                    nbits=Integer(16)
                    ),
                value=Integer(1)
                )
            )

        assert res_type6[1] == Declaration(
            Variable(
                Symbol('b'),
                type=UnsignedIntType(
                    String('uint16'),
                    nbits=Integer(16)
                    ),
                value=Integer(5)
                )
            )

        assert res_type6[2] == Declaration(
            Variable(
                Symbol('c'),
                type=UnsignedIntType(
                    String('uint16'),
                    nbits=Integer(16)
                    ),
                value=Integer(1)
                )
            )

        assert res_type6[3] == Declaration(
            Variable(
                Symbol('d'),
                type=UnsignedIntType(
                    String('uint16'),
                    nbits=Integer(16)
                    ),
                value=Integer(5)
                )
            )

        assert res_type7[0] == Declaration(
            Variable(
                Symbol('a'),
                type=UnsignedIntType(
                    String('uint32'),
                    nbits=Integer(32)
                    ),
                value=Integer(1)
                )
            )

        assert res_type7[1] == Declaration(
            Variable(
                Symbol('b'),
                type=UnsignedIntType(
                    String('uint32'),
                    nbits=Integer(32)
                    ),
                value=Integer(5)
                )
            )

        assert res_type8[0] == Declaration(
            Variable(
                Symbol('a'),
                type=UnsignedIntType(
                    String('uint64'),
                    nbits=Integer(64)
                    ),
                value=Integer(1)
                )
            )

        assert res_type8[1] == Declaration(
            Variable(
                Symbol('b'),
                type=UnsignedIntType(
                    String('uint64'),
                    nbits=Integer(64)
                    ),
                value=Integer(5)
                )
            )

        assert res_type8[2] == Declaration(
            Variable(
                Symbol('c'),
                type=UnsignedIntType(
                    String('uint64'),
                    nbits=Integer(64)
                    ),
                value=Integer(1)
                )
            )

        assert res_type8[3] == Declaration(
            Variable(
                Symbol('d'),
                type=UnsignedIntType(
                    String('uint64'),
                    nbits=Integer(64)
                    ),
                value=Integer(5)
                )
            )


    def test_float():
        c_src1 = 'float a = 1.0;'
        c_src2 = (
            'float a = 1.25;' + '\n' +
            'float b = 2.39;' + '\n'
        )
        c_src3 = 'float x = 1, y = 2;'
        c_src4 = 'float p = 5, e = 7.89;'
        c_src5 = 'float r = true, s = false;'

        # cin.TypeKind.FLOAT
        c_src_type1 = 'float x = 1, y = 2.5;'

        # cin.TypeKind.DOUBLE
        c_src_type2 = 'double x = 1, y = 2.5;'

        # cin.TypeKind.LONGDOUBLE
        c_src_type3 = 'long double x = 1, y = 2.5;'

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()
        res3 = SymPyExpression(c_src3, 'c').return_expr()
        res4 = SymPyExpression(c_src4, 'c').return_expr()
        res5 = SymPyExpression(c_src5, 'c').return_expr()

        res_type1 = SymPyExpression(c_src_type1, 'c').return_expr()
        res_type2 = SymPyExpression(c_src_type2, 'c').return_expr()
        res_type3 = SymPyExpression(c_src_type3, 'c').return_expr()

        assert res1[0] == Declaration(
            Variable(
                Symbol('a'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('1.0', precision=53)
                )
            )

        assert res2[0] == Declaration(
            Variable(
                Symbol('a'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('1.25', precision=53)
            )
        )

        assert res2[1] == Declaration(
            Variable(
                Symbol('b'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('2.3900000000000001', precision=53)
            )
        )

        assert res3[0] == Declaration(
            Variable(
                Symbol('x'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('1.0', precision=53)
            )
        )

        assert res3[1] == Declaration(
            Variable(
                Symbol('y'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('2.0', precision=53)
            )
        )

        assert res4[0] == Declaration(
            Variable(
                Symbol('p'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                ),
                value=Float('5.0', precision=53)
            )
        )

        assert res4[1] == Declaration(
            Variable(
                Symbol('e'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('7.89', precision=53)
            )
        )

        assert res5[0] == Declaration(
            Variable(
                Symbol('r'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('1.0', precision=53)
            )
        )

        assert res5[1] == Declaration(
            Variable(
                Symbol('s'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('0.0', precision=53)
            )
        )

        assert res_type1[0] == Declaration(
            Variable(
                Symbol('x'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('1.0', precision=53)
                )
            )

        assert res_type1[1] == Declaration(
            Variable(
                Symbol('y'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('2.5', precision=53)
                )
            )
        assert res_type2[0] == Declaration(
            Variable(
                Symbol('x'),
                type=FloatType(
                    String('float64'),
                    nbits=Integer(64),
                    nmant=Integer(52),
                    nexp=Integer(11)
                    ),
                value=Float('1.0', precision=53)
                )
            )

        assert res_type2[1] == Declaration(
            Variable(
                Symbol('y'),
                type=FloatType(
                    String('float64'),
                    nbits=Integer(64),
                    nmant=Integer(52),
                    nexp=Integer(11)
                    ),
                value=Float('2.5', precision=53)
                )
            )

        assert res_type3[0] == Declaration(
            Variable(
                Symbol('x'),
                type=FloatType(
                    String('float80'),
                    nbits=Integer(80),
                    nmant=Integer(63),
                    nexp=Integer(15)
                    ),
                value=Float('1.0', precision=53)
                )
            )

        assert res_type3[1] == Declaration(
            Variable(
                Symbol('y'),
                type=FloatType(
                    String('float80'),
                    nbits=Integer(80),
                    nmant=Integer(63),
                    nexp=Integer(15)
                    ),
                value=Float('2.5', precision=53)
                )
            )


    def test_bool():
        c_src1 = (
            'bool a = true, b = false;'
        )

        c_src2 = (
            'bool a = 1, b = 0;'
        )

        c_src3 = (
            'bool a = 10, b = 20;'
        )

        c_src4 = (
            'bool a = 19.1, b = 9.0, c = 0.0;'
        )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()
        res3 = SymPyExpression(c_src3, 'c').return_expr()
        res4 = SymPyExpression(c_src4, 'c').return_expr()

        assert res1[0] == Declaration(
            Variable(Symbol('a'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res1[1] == Declaration(
            Variable(Symbol('b'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res2[0] == Declaration(
            Variable(Symbol('a'),
                type=Type(String('bool')),
                value=true)
            )

        assert res2[1] == Declaration(
            Variable(Symbol('b'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res3[0] == Declaration(
            Variable(Symbol('a'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res3[1] == Declaration(
            Variable(Symbol('b'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res4[0] == Declaration(
            Variable(Symbol('a'),
                type=Type(String('bool')),
                value=true)
            )

        assert res4[1] == Declaration(
            Variable(Symbol('b'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res4[2] == Declaration(
            Variable(Symbol('c'),
                type=Type(String('bool')),
                value=false
                )
            )

    @XFAIL # this is expected to fail because of a bug in the C parser.
    def test_function():
        c_src1 = (
            'void fun1()' + '\n' +
            '{' + '\n' +
            'int a;' + '\n' +
            '}'
        )
        c_src2 = (
            'int fun2()' + '\n' +
            '{'+ '\n' +
            'int a;' + '\n' +
            'return a;' + '\n' +
            '}'
        )
        c_src3 = (
            'float fun3()' + '\n' +
            '{' + '\n' +
            'float b;' + '\n' +
            'return b;' + '\n' +
            '}'
        )
        c_src4 = (
            'float fun4()' + '\n' +
            '{}'
        )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()
        res3 = SymPyExpression(c_src3, 'c').return_expr()
        res4 = SymPyExpression(c_src4, 'c').return_expr()

        assert res1[0] == FunctionDefinition(
            NoneToken(),
            name=String('fun1'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(
                        Symbol('a'),
                        type=IntBaseType(String('intc'))
                    )
                )
            )
        )

        assert res2[0] == FunctionDefinition(
            IntBaseType(String('intc')),
            name=String('fun2'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(
                        Symbol('a'),
                        type=IntBaseType(String('intc'))
                    )
                ),
                Return('a')
            )
        )

        assert res3[0] == FunctionDefinition(
            FloatType(
                String('float32'),
                nbits=Integer(32),
                nmant=Integer(23),
                nexp=Integer(8)
                ),
            name=String('fun3'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(
                        Symbol('b'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            )
                    )
                ),
                Return('b')
            )
        )

        assert res4[0] == FunctionPrototype(
            FloatType(
                String('float32'),
                nbits=Integer(32),
                nmant=Integer(23),
                nexp=Integer(8)
                ),
            name=String('fun4'),
            parameters=()
        )

    @XFAIL # this is expected to fail because of a bug in the C parser.
    def test_parameters():
        c_src1 = (
            'void fun1( int a)' + '\n' +
            '{' + '\n' +
            'int i;' + '\n' +
            '}'
        )
        c_src2 = (
            'int fun2(float x, float y)' + '\n' +
            '{'+ '\n' +
            'int a;' + '\n' +
            'return a;' + '\n' +
            '}'
        )
        c_src3 = (
            'float fun3(int p, float q, int r)' + '\n' +
            '{' + '\n' +
            'float b;' + '\n' +
            'return b;' + '\n' +
            '}'
        )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()
        res3 = SymPyExpression(c_src3, 'c').return_expr()

        assert res1[0] == FunctionDefinition(
            NoneToken(),
            name=String('fun1'),
            parameters=(
                Variable(
                    Symbol('a'),
                    type=IntBaseType(String('intc'))
                ),
            ),
            body=CodeBlock(
                Declaration(
                    Variable(
                        Symbol('i'),
                        type=IntBaseType(String('intc'))
                    )
                )
            )
        )

        assert res2[0] == FunctionDefinition(
            IntBaseType(String('intc')),
            name=String('fun2'),
            parameters=(
                Variable(
                    Symbol('x'),
                    type=FloatType(
                        String('float32'),
                        nbits=Integer(32),
                        nmant=Integer(23),
                        nexp=Integer(8)
                        )
                ),
                Variable(
                    Symbol('y'),
                    type=FloatType(
                        String('float32'),
                        nbits=Integer(32),
                        nmant=Integer(23),
                        nexp=Integer(8)
                        )
                )
            ),
            body=CodeBlock(
                Declaration(
                    Variable(
                        Symbol('a'),
                        type=IntBaseType(String('intc'))
                    )
                ),
                Return('a')
            )
        )

        assert res3[0] == FunctionDefinition(
            FloatType(
                String('float32'),
                nbits=Integer(32),
                nmant=Integer(23),
                nexp=Integer(8)
                ),
            name=String('fun3'),
            parameters=(
                Variable(
                    Symbol('p'),
                    type=IntBaseType(String('intc'))
                ),
                Variable(
                    Symbol('q'),
                    type=FloatType(
                        String('float32'),
                        nbits=Integer(32),
                        nmant=Integer(23),
                        nexp=Integer(8)
                        )
                ),
                Variable(
                    Symbol('r'),
                    type=IntBaseType(String('intc'))
                )
            ),
            body=CodeBlock(
                Declaration(
                    Variable(
                        Symbol('b'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            )
                    )
                ),
                Return('b')
            )
        )

    @XFAIL # this is expected to fail because of a bug in the C parser.
    def test_function_call():
        c_src1 = (
            'int fun1(int x)' + '\n' +
            '{' + '\n' +
            'return x;' + '\n' +
            '}' + '\n' +
            'void caller()' + '\n' +
            '{' + '\n' +
            'int x = fun1(2);' + '\n' +
            '}'
        )

        c_src2 = (
            'int fun2(int a, int b, int c)' + '\n' +
            '{' + '\n' +
            'return a;' + '\n' +
            '}' + '\n' +
            'void caller()' + '\n' +
            '{' + '\n' +
            'int y = fun2(2, 3, 4);' + '\n' +
            '}'
        )

        c_src3 = (
            'int fun3(int a, int b, int c)' + '\n' +
            '{' + '\n' +
            'return b;' + '\n' +
            '}' + '\n' +
            'void caller()' + '\n' +
            '{' + '\n' +
            'int p;' + '\n' +
            'int q;' + '\n' +
            'int r;' + '\n' +
            'int z = fun3(p, q, r);' + '\n' +
            '}'
        )

        c_src4 = (
            'int fun4(float a, float b, int c)' + '\n' +
            '{' + '\n' +
            'return c;' + '\n' +
            '}' + '\n' +
            'void caller()' + '\n' +
            '{' + '\n' +
            'float x;' + '\n' +
            'float y;' + '\n' +
            'int z;' + '\n' +
            'int i = fun4(x, y, z)' + '\n' +
            '}'
        )

        c_src5 = (
            'int fun()' + '\n' +
            '{' + '\n' +
            'return 1;' + '\n' +
            '}' + '\n' +
            'void caller()' + '\n' +
            '{' + '\n' +
            'int a = fun()' + '\n' +
            '}'
        )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()
        res3 = SymPyExpression(c_src3, 'c').return_expr()
        res4 = SymPyExpression(c_src4, 'c').return_expr()
        res5 = SymPyExpression(c_src5, 'c').return_expr()


        assert res1[0] == FunctionDefinition(
            IntBaseType(String('intc')),
            name=String('fun1'),
            parameters=(Variable(Symbol('x'),
                type=IntBaseType(String('intc'))
                ),
            ),
            body=CodeBlock(
                Return('x')
                )
            )

        assert res1[1] == FunctionDefinition(
            NoneToken(),
            name=String('caller'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('x'),
                        value=FunctionCall(String('fun1'),
                            function_args=(
                                Integer(2),
                                )
                            )
                        )
                    )
                )
            )

        assert res2[0] == FunctionDefinition(
            IntBaseType(String('intc')),
            name=String('fun2'),
            parameters=(Variable(Symbol('a'),
                type=IntBaseType(String('intc'))
                ),
            Variable(Symbol('b'),
                type=IntBaseType(String('intc'))
                ),
            Variable(Symbol('c'),
                type=IntBaseType(String('intc'))
                )
            ),
            body=CodeBlock(
                Return('a')
                )
            )

        assert res2[1] == FunctionDefinition(
            NoneToken(),
            name=String('caller'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('y'),
                        value=FunctionCall(
                            String('fun2'),
                            function_args=(
                                Integer(2),
                                Integer(3),
                                Integer(4)
                                )
                            )
                        )
                    )
                )
            )

        assert res3[0] == FunctionDefinition(
            IntBaseType(String('intc')),
            name=String('fun3'),
            parameters=(
                Variable(Symbol('a'),
                    type=IntBaseType(String('intc'))
                    ),
                Variable(Symbol('b'),
                    type=IntBaseType(String('intc'))
                    ),
                Variable(Symbol('c'),
                    type=IntBaseType(String('intc'))
                    )
                ),
            body=CodeBlock(
                Return('b')
                )
            )

        assert res3[1] == FunctionDefinition(
            NoneToken(),
            name=String('caller'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('p'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('q'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('r'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('z'),
                        value=FunctionCall(
                            String('fun3'),
                            function_args=(
                                Symbol('p'),
                                Symbol('q'),
                                Symbol('r')
                                )
                            )
                        )
                    )
                )
            )

        assert res4[0] == FunctionDefinition(
            IntBaseType(String('intc')),
            name=String('fun4'),
            parameters=(Variable(Symbol('a'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    )
                ),
            Variable(Symbol('b'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    )
                ),
            Variable(Symbol('c'),
                type=IntBaseType(String('intc'))
                )
            ),
            body=CodeBlock(
                Return('c')
                )
            )

        assert res4[1] == FunctionDefinition(
            NoneToken(),
            name=String('caller'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('x'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            )
                        )
                    ),
                Declaration(
                    Variable(Symbol('y'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            )
                        )
                    ),
                Declaration(
                    Variable(Symbol('z'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('i'),
                        value=FunctionCall(String('fun4'),
                            function_args=(
                                Symbol('x'),
                                Symbol('y'),
                                Symbol('z')
                                )
                            )
                        )
                    )
                )
            )

        assert res5[0] == FunctionDefinition(
            IntBaseType(String('intc')),
            name=String('fun'),
            parameters=(),
            body=CodeBlock(
                Return('')
                )
            )

        assert res5[1] == FunctionDefinition(
            NoneToken(),
            name=String('caller'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        value=FunctionCall(String('fun'),
                            function_args=()
                            )
                        )
                    )
                )
            )


    def test_parse():
        c_src1 = (
            'int a;' + '\n' +
            'int b;' + '\n'
        )
        c_src2 = (
            'void fun1()' + '\n' +
            '{' + '\n' +
            'int a;' + '\n' +
            '}'
        )

        f1 = open('..a.h', 'w')
        f2 = open('..b.h', 'w')

        f1.write(c_src1)
        f2. write(c_src2)

        f1.close()
        f2.close()

        res1 = SymPyExpression('..a.h', 'c').return_expr()
        res2 = SymPyExpression('..b.h', 'c').return_expr()

        os.remove('..a.h')
        os.remove('..b.h')

        assert res1[0] == Declaration(
            Variable(
                Symbol('a'),
                type=IntBaseType(String('intc'))
            )
        )
        assert res1[1] == Declaration(
            Variable(
                Symbol('b'),
                type=IntBaseType(String('intc'))
            )
        )
        assert res2[0] == FunctionDefinition(
            NoneToken(),
            name=String('fun1'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(
                        Symbol('a'),
                        type=IntBaseType(String('intc'))
                    )
                )
            )
        )


    def test_binary_operators():
        c_src1 = (
            'void func()'+
            '{' + '\n' +
                'int a;' + '\n' +
                'a = 1;' + '\n' +
            '}'
        )
        c_src2 = (
            'void func()'+
            '{' + '\n' +
                'int a = 0;' + '\n' +
                'a = a + 1;' + '\n' +
                'a = 3*a - 10;' + '\n' +
            '}'
        )
        c_src3 = (
            'void func()'+
            '{' + '\n' +
                'int a = 10;' + '\n' +
                'a = 1 + a - 3 * 6;' + '\n' +
            '}'
        )
        c_src4 = (
            'void func()'+
            '{' + '\n' +
                'int a;' + '\n' +
                'int b;' + '\n' +
                'a = 100;' + '\n' +
                'b = a*a + a*a + a + 19*a + 1 + 24;' + '\n' +
            '}'
        )
        c_src5 = (
            'void func()'+
            '{' + '\n' +
                'int a;' + '\n' +
                'int b;' + '\n' +
                'int c;' + '\n' +
                'int d;' + '\n' +
                'a = 1;' + '\n' +
                'b = 2;' + '\n' +
                'c = b;' + '\n' +
                'd = ((a+b)*(a+c))*((c-d)*(a+c));' + '\n' +
            '}'
        )
        c_src6 = (
            'void func()'+
            '{' + '\n' +
                'int a;' + '\n' +
                'int b;' + '\n' +
                'int c;' + '\n' +
                'int d;' + '\n' +
                'a = 1;' + '\n' +
                'b = 2;' + '\n' +
                'c = 3;' + '\n' +
                'd = (a*a*a*a + 3*b*b + b + b + c*d);' + '\n' +
            '}'
        )
        c_src7 = (
            'void func()'+
            '{' + '\n' +
                'float a;' + '\n' +
                'a = 1.01;' + '\n' +
            '}'
        )

        c_src8 = (
            'void func()'+
            '{' + '\n' +
                'float a;' + '\n' +
                'a = 10.0 + 2.5;' + '\n' +
            '}'
        )

        c_src9 = (
            'void func()'+
            '{' + '\n' +
                'float a;' + '\n' +
                'a = 10.0 / 2.5;' + '\n' +
            '}'
        )

        c_src10 = (
            'void func()'+
            '{' + '\n' +
                'int a;' + '\n' +
                'a = 100 / 4;' + '\n' +
            '}'
        )

        c_src11 = (
            'void func()'+
            '{' + '\n' +
                'int a;' + '\n' +
                'a = 20 - 100 / 4 * 5 + 10;' + '\n' +
            '}'
        )

        c_src12 = (
            'void func()'+
            '{' + '\n' +
                'int a;' + '\n' +
                'a = (20 - 100) / 4 * (5 + 10);' + '\n' +
            '}'
        )

        c_src13 = (
            'void func()'+
            '{' + '\n' +
                'int a;' + '\n' +
                'int b;' + '\n' +
                'float c;' + '\n' +
                'c = b/a;' + '\n' +
            '}'
        )

        c_src14 = (
            'void func()'+
            '{' + '\n' +
                'int a = 2;' + '\n' +
                'int d = 5;' + '\n' +
                'int n = 10;' + '\n' +
                'int s;' + '\n' +
                's = (a/2)*(2*a + (n-1)*d);' + '\n' +
            '}'
        )

        c_src15 = (
            'void func()'+
            '{' + '\n' +
                'int a;' + '\n' +
                'a = 1 % 2;' + '\n' +
            '}'
        )

        c_src16 = (
            'void func()'+
            '{' + '\n' +
                'int a = 2;' + '\n' +
                'int b;' + '\n' +
                'b = a % 3;' + '\n' +
            '}'
        )

        c_src17 = (
            'void func()'+
            '{' + '\n' +
                'int a = 100;' + '\n' +
                'int b = 3;' + '\n' +
                'int c;' + '\n' +
                'c = a % b;' + '\n' +
            '}'
        )

        c_src18 = (
            'void func()'+
            '{' + '\n' +
                'int a = 100;' + '\n' +
                'int b = 3;' + '\n' +
                'int mod = 1000000007;' + '\n' +
                'int c;' + '\n' +
                'c = (a + b * (100/a)) % mod;' + '\n' +
            '}'
        )

        c_src19 = (
            'void func()'+
            '{' + '\n' +
                'int a = 100;' + '\n' +
                'int b = 3;' + '\n' +
                'int mod = 1000000007;' + '\n' +
                'int c;' + '\n' +
                'c = ((a % mod + b % mod) % mod' \
                '* (a % mod - b % mod) % mod) % mod;' + '\n' +
            '}'
        )

        c_src20 = (
            'void func()'+
            '{' + '\n' +
            'bool a' + '\n' +
            'bool b;' + '\n' +
            'a = 1 == 2;' + '\n' +
            'b = 1 != 2;' + '\n' +
            '}'
        )

        c_src21 = (
            'void func()'+
            '{' + '\n' +
            'bool a;' + '\n' +
            'bool b;' + '\n' +
            'bool c;' + '\n' +
            'bool d;' + '\n' +
            'a = 1 == 2;' + '\n' +
            'b = 1 <= 2;' + '\n' +
            'c = 1 > 2;' + '\n' +
            'd = 1 >= 2;' + '\n' +
            '}'
        )

        c_src22 = (
            'void func()'+
            '{' + '\n' +
            'int a = 1;' + '\n' +
            'int b = 2;' + '\n' +

            'bool c1;' + '\n' +
            'bool c2;' + '\n' +
            'bool c3;' + '\n' +
            'bool c4;' + '\n' +
            'bool c5;' + '\n' +
            'bool c6;' + '\n' +
            'bool c7;' + '\n' +
            'bool c8;' + '\n' +

            'c1 = a == 1;' + '\n' +
            'c2 = b == 2;' + '\n' +

            'c3 = 1 != a;' + '\n' +
            'c4 = 1 != b;' + '\n' +

            'c5 = a < 0;' + '\n' +
            'c6 = b <= 10;' + '\n' +
            'c7 = a > 0;' + '\n' +
            'c8 = b >= 11;' + '\n' +
            '}'
        )

        c_src23 = (
            'void func()'+
            '{' + '\n' +
            'int a = 3;' + '\n' +
            'int b = 4;' + '\n' +

            'bool c1;' + '\n' +
            'bool c2;' + '\n' +
            'bool c3;' + '\n' +
            'bool c4;' + '\n' +
            'bool c5;' + '\n' +
            'bool c6;' + '\n' +

            'c1 = a == b;' + '\n' +
            'c2 = a != b;' + '\n' +
            'c3 = a < b;' + '\n' +
            'c4 = a <= b;' + '\n' +
            'c5 = a > b;' + '\n' +
            'c6 = a >= b;' + '\n' +
            '}'
        )

        c_src24 = (
            'void func()'+
            '{' + '\n' +
            'float a = 1.25'
            'float b = 2.5;' + '\n' +

            'bool c1;' + '\n' +
            'bool c2;' + '\n' +
            'bool c3;' + '\n' +
            'bool c4;' + '\n' +

            'c1 = a == 1.25;' + '\n' +
            'c2 = b == 2.54;' + '\n' +

            'c3 = 1.2 != a;' + '\n' +
            'c4 = 1.5 != b;' + '\n' +
            '}'
        )

        c_src25 = (
            'void func()'+
            '{' + '\n' +
            'float a = 1.25' + '\n' +
            'float b = 2.5;' + '\n' +

            'bool c1;' + '\n' +
            'bool c2;' + '\n' +
            'bool c3;' + '\n' +
            'bool c4;' + '\n' +
            'bool c5;' + '\n' +
            'bool c6;' + '\n' +

            'c1 = a == b;' + '\n' +
            'c2 = a != b;' + '\n' +
            'c3 = a < b;' + '\n' +
            'c4 = a <= b;' + '\n' +
            'c5 = a > b;' + '\n' +
            'c6 = a >= b;' + '\n' +
            '}'
        )

        c_src26 = (
            'void func()'+
            '{' + '\n' +
            'bool c1;' + '\n' +
            'bool c2;' + '\n' +
            'bool c3;' + '\n' +
            'bool c4;' + '\n' +
            'bool c5;' + '\n' +
            'bool c6;' + '\n' +

            'c1 = true == true;' + '\n' +
            'c2 = true == false;' + '\n' +
            'c3 = false == false;' + '\n' +

            'c4 = true != true;' + '\n' +
            'c5 = true != false;' + '\n' +
            'c6 = false != false;' + '\n' +
            '}'
        )

        c_src27 = (
            'void func()'+
            '{' + '\n' +
            'bool c1;' + '\n' +
            'bool c2;' + '\n' +
            'bool c3;' + '\n' +
            'bool c4;' + '\n' +
            'bool c5;' + '\n' +
            'bool c6;' + '\n' +

            'c1 = true && true;' + '\n' +
            'c2 = true && false;' + '\n' +
            'c3 = false && false;' + '\n' +

            'c4 = true || true;' + '\n' +
            'c5 = true || false;' + '\n' +
            'c6 = false || false;' + '\n' +
            '}'
        )

        c_src28 = (
            'void func()'+
            '{' + '\n' +
            'bool a;' + '\n' +
            'bool c1;' + '\n' +
            'bool c2;' + '\n' +
            'bool c3;' + '\n' +
            'bool c4;' + '\n' +

            'c1 = a && true;' + '\n' +
            'c2 = false && a;' + '\n' +

            'c3 = true || a;' + '\n' +
            'c4 = a || false;' + '\n' +
            '}'
        )

        c_src29 = (
            'void func()'+
            '{' + '\n' +
            'int a;' + '\n' +
            'bool c1;' + '\n' +
            'bool c2;' + '\n' +
            'bool c3;' + '\n' +
            'bool c4;' + '\n' +

            'c1 = a && 1;' + '\n' +
            'c2 = a && 0;' + '\n' +

            'c3 = a || 1;' + '\n' +
            'c4 = 0 || a;' + '\n' +
            '}'
        )

        c_src30 = (
            'void func()'+
            '{' + '\n' +
            'int a;' + '\n' +
            'int b;' + '\n' +
            'bool c;'+ '\n' +
            'bool d;'+ '\n' +

            'bool c1;' + '\n' +
            'bool c2;' + '\n' +
            'bool c3;' + '\n' +
            'bool c4;' + '\n' +
            'bool c5;' + '\n' +
            'bool c6;' + '\n' +

            'c1 = a && b;' + '\n' +
            'c2 = a && c;' + '\n' +
            'c3 = c && d;' + '\n' +

            'c4 = a || b;' + '\n' +
            'c5 = a || c;' + '\n' +
            'c6 = c || d;' + '\n' +
            '}'
        )

        c_src_raise1 = (
            'void func()'+
            '{' + '\n' +
            'int a;' + '\n' +
            'a = -1;' + '\n' +
            '}'
        )

        c_src_raise2 = (
            'void func()'+
            '{' + '\n' +
            'int a;' + '\n' +
            'a = -+1;' + '\n' +
            '}'
        )

        c_src_raise3 = (
            'void func()'+
            '{' + '\n' +
            'int a;' + '\n' +
            'a = 2*-2;' + '\n' +
            '}'
        )

        c_src_raise4 = (
            'void func()'+
            '{' + '\n' +
            'int a;' + '\n' +
            'a = (int)2.0;' + '\n' +
            '}'
        )

        c_src_raise5 = (
            'void func()'+
            '{' + '\n' +
            'int a=100;' + '\n' +
            'a = (a==100)?(1):(0);' + '\n' +
            '}'
        )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()
        res3 = SymPyExpression(c_src3, 'c').return_expr()
        res4 = SymPyExpression(c_src4, 'c').return_expr()
        res5 = SymPyExpression(c_src5, 'c').return_expr()
        res6 = SymPyExpression(c_src6, 'c').return_expr()
        res7 = SymPyExpression(c_src7, 'c').return_expr()
        res8 = SymPyExpression(c_src8, 'c').return_expr()
        res9 = SymPyExpression(c_src9, 'c').return_expr()
        res10 = SymPyExpression(c_src10, 'c').return_expr()
        res11 = SymPyExpression(c_src11, 'c').return_expr()
        res12 = SymPyExpression(c_src12, 'c').return_expr()
        res13 = SymPyExpression(c_src13, 'c').return_expr()
        res14 = SymPyExpression(c_src14, 'c').return_expr()
        res15 = SymPyExpression(c_src15, 'c').return_expr()
        res16 = SymPyExpression(c_src16, 'c').return_expr()
        res17 = SymPyExpression(c_src17, 'c').return_expr()
        res18 = SymPyExpression(c_src18, 'c').return_expr()
        res19 = SymPyExpression(c_src19, 'c').return_expr()
        res20 = SymPyExpression(c_src20, 'c').return_expr()
        res21 = SymPyExpression(c_src21, 'c').return_expr()
        res22 = SymPyExpression(c_src22, 'c').return_expr()
        res23 = SymPyExpression(c_src23, 'c').return_expr()
        res24 = SymPyExpression(c_src24, 'c').return_expr()
        res25 = SymPyExpression(c_src25, 'c').return_expr()
        res26 = SymPyExpression(c_src26, 'c').return_expr()
        res27 = SymPyExpression(c_src27, 'c').return_expr()
        res28 = SymPyExpression(c_src28, 'c').return_expr()
        res29 = SymPyExpression(c_src29, 'c').return_expr()
        res30 = SymPyExpression(c_src30, 'c').return_expr()

        assert res1[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(Variable(Symbol('a')), Integer(1))
                )
            )

        assert res2[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                    type=IntBaseType(String('intc')),
                    value=Integer(0))),
                Assignment(
                    Variable(Symbol('a')),
                    Add(Symbol('a'),
                        Integer(1))
                    ),
                Assignment(Variable(Symbol('a')),
                    Add(
                        Mul(
                            Integer(3),
                            Symbol('a')),
                        Integer(-10)
                        )
                    )
                )
            )

        assert res3[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(10)
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Add(
                        Symbol('a'),
                        Integer(-17)
                        )
                    )
                )
            )

        assert res4[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Integer(100)),
                Assignment(
                    Variable(Symbol('b')),
                    Add(
                        Mul(
                            Integer(2),
                        Pow(
                            Symbol('a'),
                            Integer(2))
                        ),
                        Mul(
                            Integer(20),
                            Symbol('a')),
                        Integer(25)
                        )
                    )
                )
            )

        assert res5[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('d'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Integer(1)),
                Assignment(
                    Variable(Symbol('b')),
                    Integer(2)
                    ),
                Assignment(
                    Variable(Symbol('c')),
                    Symbol('b')),
                Assignment(
                    Variable(Symbol('d')),
                    Mul(
                        Add(
                            Symbol('a'),
                            Symbol('b')),
                        Pow(
                            Add(
                                Symbol('a'),
                                Symbol('c')
                                ),
                            Integer(2)
                            ),
                        Add(
                            Symbol('c'),
                            Mul(
                                Integer(-1),
                                Symbol('d')
                                )
                            )
                        )
                    )
                )
            )

        assert res6[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('d'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Integer(1)
                    ),
                Assignment(
                    Variable(Symbol('b')),
                    Integer(2)
                    ),
                Assignment(
                    Variable(Symbol('c')),
                    Integer(3)
                    ),
                Assignment(
                    Variable(Symbol('d')),
                    Add(
                        Pow(
                            Symbol('a'),
                            Integer(4)
                            ),
                        Mul(
                            Integer(3),
                            Pow(
                                Symbol('b'),
                                Integer(2)
                                )
                            ),
                        Mul(
                            Integer(2),
                            Symbol('b')
                            ),
                        Mul(
                            Symbol('c'),
                            Symbol('d')
                            )
                        )
                    )
                )
            )

        assert res7[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            )
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Float('1.01', precision=53)
                    )
                )
            )

        assert res8[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            )
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Float('12.5', precision=53)
                    )
                )
            )

        assert res9[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            )
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Float('4.0', precision=53)
                    )
                )
            )

        assert res10[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Integer(25)
                    )
                )
            )

        assert res11[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Integer(-95)
                    )
                )
            )

        assert res12[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Integer(-300)
                    )
                )
            )

        assert res13[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            )
                        )
                    ),
                Assignment(
                    Variable(Symbol('c')),
                    Mul(
                        Pow(
                            Symbol('a'),
                            Integer(-1)
                            ),
                        Symbol('b')
                        )
                    )
                )
            )

        assert res14[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(2)
                        )
                    ),
                Declaration(
                    Variable(Symbol('d'),
                        type=IntBaseType(String('intc')),
                        value=Integer(5)
                        )
                    ),
                Declaration(
                    Variable(Symbol('n'),
                        type=IntBaseType(String('intc')),
                        value=Integer(10)
                        )
                    ),
                Declaration(
                    Variable(Symbol('s'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('s')),
                    Mul(
                        Rational(1, 2),
                        Symbol('a'),
                        Add(
                            Mul(
                                Integer(2),
                                Symbol('a')
                                ),
                            Mul(
                                Symbol('d'),
                                Add(
                                    Symbol('n'),
                                    Integer(-1)
                                    )
                                )
                            )
                        )
                    )
                )
            )

        assert res15[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Integer(1)
                    )
                )
            )

        assert res16[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(2)
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('b')),
                    Mod(
                        Symbol('a'),
                        Integer(3)
                        )
                    )
                )
            )

        assert res17[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(100)
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc')),
                        value=Integer(3)
                        )
                    ),
                Declaration(
                    Variable(Symbol('c'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c')),
                    Mod(
                        Symbol('a'),
                        Symbol('b')
                        )
                    )
                )
            )

        assert res18[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(100)
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc')),
                        value=Integer(3)
                        )
                    ),
                Declaration(
                    Variable(Symbol('mod'),
                        type=IntBaseType(String('intc')),
                        value=Integer(1000000007)
                        )
                    ),
                Declaration(
                    Variable(Symbol('c'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c')),
                    Mod(
                        Add(
                            Symbol('a'),
                            Mul(
                                Integer(100),
                                Pow(
                                    Symbol('a'),
                                    Integer(-1)
                                    ),
                                Symbol('b')
                                )
                            ),
                        Symbol('mod')
                        )
                    )
                )
            )

        assert res19[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(100)
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc')),
                        value=Integer(3)
                        )
                    ),
                Declaration(
                    Variable(Symbol('mod'),
                        type=IntBaseType(String('intc')),
                        value=Integer(1000000007)
                        )
                    ),
                Declaration(
                    Variable(Symbol('c'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c')),
                    Mod(
                        Mul(
                            Add(
                                Mod(
                                    Symbol('a'),
                                    Symbol('mod')
                                    ),
                                Mul(
                                    Integer(-1),
                                    Mod(
                                        Symbol('b'),
                                        Symbol('mod')
                                        )
                                    )
                                ),
                            Mod(
                                Add(
                                    Symbol('a'),
                                    Symbol('b')
                                    ),
                                Symbol('mod')
                                )
                            ),
                        Symbol('mod')
                        )
                    )
                )
            )

        assert res20[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    false
                    ),
                Assignment(
                    Variable(Symbol('b')),
                    true
                    )
                )
            )

        assert res21[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('d'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    false
                    ),
                Assignment(
                    Variable(Symbol('b')),
                    true
                    ),
                Assignment(
                    Variable(Symbol('c')),
                    false
                    ),
                Assignment(
                    Variable(Symbol('d')),
                    false
                    )
                )
            )

        assert res22[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(1)
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc')),
                        value=Integer(2)
                        )
                    ),
                Declaration(
                    Variable(Symbol('c1'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c2'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c3'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c4'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c5'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c6'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c7'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c8'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c1')),
                    Equality(
                        Symbol('a'),
                        Integer(1)
                        )
                    ),
                Assignment(
                    Variable(Symbol('c2')),
                    Equality(
                        Symbol('b'),
                        Integer(2)
                        )
                    ),
                Assignment(
                    Variable(Symbol('c3')),
                    Unequality(
                        Integer(1),
                        Symbol('a')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c4')),
                    Unequality(
                        Integer(1),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c5')),
                    StrictLessThan(
                        Symbol('a'),
                        Integer(0)
                        )
                    ),
                Assignment(
                    Variable(Symbol('c6')),
                    LessThan(
                        Symbol('b'),
                        Integer(10)
                        )
                    ),
                Assignment(
                    Variable(Symbol('c7')),
                    StrictGreaterThan(
                        Symbol('a'),
                        Integer(0)
                        )
                    ),
                Assignment(
                    Variable(Symbol('c8')),
                    GreaterThan(
                        Symbol('b'),
                        Integer(11)
                        )
                    )
                )
            )

        assert res23[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(3)
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc')),
                        value=Integer(4)
                        )
                    ),
                Declaration(
                    Variable(Symbol('c1'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c2'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c3'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c4'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c5'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c6'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c1')),
                    Equality(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c2')),
                    Unequality(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c3')),
                    StrictLessThan(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c4')),
                    LessThan(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c5')),
                    StrictGreaterThan(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c6')),
                    GreaterThan(
                        Symbol('a'),
                        Symbol('b')
                        )
                    )
                )
            )

        assert res24[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            )
                        )
                    ),
                Declaration(
                    Variable(Symbol('c1'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c2'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c3'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c4'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c1')),
                    Equality(
                        Symbol('a'),
                        Float('1.25', precision=53)
                        )
                    ),
                Assignment(
                    Variable(Symbol('c3')),
                    Unequality(
                        Float('1.2', precision=53),
                        Symbol('a')
                        )
                    )
                )
            )


        assert res25[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            ),
                        value=Float('1.25', precision=53)
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=FloatType(
                            String('float32'),
                            nbits=Integer(32),
                            nmant=Integer(23),
                            nexp=Integer(8)
                            ),
                        value=Float('2.5', precision=53)
                        )
                    ),
                Declaration(
                    Variable(Symbol('c1'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c2'),
                        type=Type(String('bool')
                            )
                        )
                    ),
                Declaration(
                    Variable(Symbol('c3'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c4'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c5'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c6'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c1')),
                    Equality(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c2')),
                    Unequality(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c3')),
                    StrictLessThan(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c4')),
                    LessThan(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c5')),
                    StrictGreaterThan(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c6')),
                    GreaterThan(
                        Symbol('a'),
                        Symbol('b')
                        )
                    )
                )
            )

        assert res26[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(), body=CodeBlock(
                Declaration(
                    Variable(Symbol('c1'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c2'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c3'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c4'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c5'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c6'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c1')),
                    true
                    ),
                Assignment(
                    Variable(Symbol('c2')),
                    false
                    ),
                Assignment(
                    Variable(Symbol('c3')),
                    true
                    ),
                Assignment(
                    Variable(Symbol('c4')),
                    false
                    ),
                Assignment(
                    Variable(Symbol('c5')),
                    true
                    ),
                Assignment(
                    Variable(Symbol('c6')),
                    false
                    )
                )
            )

        assert res27[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('c1'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c2'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c3'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c4'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c5'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c6'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c1')),
                    true
                    ),
                Assignment(
                    Variable(Symbol('c2')),
                    false
                    ),
                Assignment(
                    Variable(Symbol('c3')),
                    false
                    ),
                Assignment(
                    Variable(Symbol('c4')),
                    true
                    ),
                Assignment(
                    Variable(Symbol('c5')),
                    true
                    ),
                Assignment(
                    Variable(Symbol('c6')),
                    false)
                )
            )

        assert res28[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c1'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c2'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c3'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c4'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c1')),
                    Symbol('a')
                    ),
                Assignment(
                    Variable(Symbol('c2')),
                    false
                    ),
                Assignment(
                    Variable(Symbol('c3')),
                    true
                    ),
                Assignment(
                    Variable(Symbol('c4')),
                    Symbol('a')
                    )
                )
            )

        assert res29[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c1'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c2'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c3'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c4'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c1')),
                    Symbol('a')
                    ),
                Assignment(
                    Variable(Symbol('c2')),
                    false
                    ),
                Assignment(
                    Variable(Symbol('c3')),
                    true
                    ),
                Assignment(
                    Variable(Symbol('c4')),
                    Symbol('a')
                    )
                )
            )

        assert res30[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('d'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c1'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c2'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c3'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c4'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c5'),
                        type=Type(String('bool'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('c6'),
                        type=Type(String('bool'))
                        )
                    ),
                Assignment(
                    Variable(Symbol('c1')),
                    And(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c2')),
                    And(
                        Symbol('a'),
                        Symbol('c')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c3')),
                    And(
                        Symbol('c'),
                        Symbol('d')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c4')),
                    Or(
                        Symbol('a'),
                        Symbol('b')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c5')),
                    Or(
                        Symbol('a'),
                        Symbol('c')
                        )
                    ),
                Assignment(
                    Variable(Symbol('c6')),
                    Or(
                        Symbol('c'),
                        Symbol('d')
                        )
                    )
                )
            )

        raises(NotImplementedError, lambda: SymPyExpression(c_src_raise1, 'c'))
        raises(NotImplementedError, lambda: SymPyExpression(c_src_raise2, 'c'))
        raises(NotImplementedError, lambda: SymPyExpression(c_src_raise3, 'c'))
        raises(NotImplementedError, lambda: SymPyExpression(c_src_raise4, 'c'))
        raises(NotImplementedError, lambda: SymPyExpression(c_src_raise5, 'c'))


    @XFAIL
    def test_var_decl():
        c_src1 = (
            'int b = 100;' + '\n' +
            'int a = b;' + '\n'
        )

        c_src2 = (
            'int a = 1;' + '\n' +
            'int b = a + 1;' + '\n'
        )

        c_src3 = (
            'float a = 10.0 + 2.5;' + '\n' +
            'float b = a * 20.0;' + '\n'
        )

        c_src4 = (
            'int a = 1 + 100 - 3 * 6;' + '\n'
        )

        c_src5 = (
            'int a = (((1 + 100) * 12) - 3) * (6 - 10);' + '\n'
        )

        c_src6 = (
            'int b = 2;' + '\n' +
            'int c = 3;' + '\n' +
            'int a = b + c * 4;' + '\n'
        )

        c_src7 = (
            'int b = 1;' + '\n' +
            'int c = b + 2;' + '\n' +
            'int a = 10 * b * b * c;' + '\n'
        )

        c_src8 = (
            'void func()'+
            '{' + '\n' +
                'int a = 1;' + '\n' +
                'int b = 2;' + '\n' +
                'int temp = a;' + '\n' +
                'a = b;' + '\n' +
                'b = temp;' + '\n' +
            '}'
        )

        c_src9 = (
            'int a = 1;' + '\n' +
            'int b = 2;' + '\n' +
            'int c = a;' + '\n' +
            'int d = a + b + c;' + '\n' +
            'int e = a*a*a + 3*a*a*b + 3*a*b*b + b*b*b;' + '\n'
            'int f = (a + b + c) * (a + b - c);' + '\n' +
            'int g = (a + b + c + d)*(a + b + c + d)*(a * (b - c));'
            + '\n'
        )

        c_src10 = (
            'float a = 10.0;' + '\n' +
            'float b = 2.5;' + '\n' +
            'float c = a*a + 2*a*b + b*b;' + '\n'
        )

        c_src11 = (
            'float a = 10.0 / 2.5;' + '\n'
        )

        c_src12 = (
            'int a = 100 / 4;' + '\n'
        )

        c_src13 = (
            'int a = 20 - 100 / 4 * 5 + 10;' + '\n'
        )

        c_src14 = (
            'int a = (20 - 100) / 4 * (5 + 10);' + '\n'
        )

        c_src15 = (
            'int a = 4;' + '\n' +
            'int b = 2;' + '\n' +
            'float c = b/a;' + '\n'
        )

        c_src16 = (
            'int a = 2;' + '\n' +
            'int d = 5;' + '\n' +
            'int n = 10;' + '\n' +
            'int s = (a/2)*(2*a + (n-1)*d);' + '\n'
        )

        c_src17 = (
            'int a = 1 % 2;' + '\n'
        )

        c_src18 = (
            'int a = 2;' + '\n' +
            'int b = a % 3;' + '\n'
        )

        c_src19 = (
            'int a = 100;' + '\n' +
            'int b = 3;' + '\n' +
            'int c = a % b;' + '\n'
        )

        c_src20 = (
            'int a = 100;' + '\n' +
            'int b = 3;' + '\n' +
            'int mod = 1000000007;' + '\n' +
            'int c = (a + b * (100/a)) % mod;' + '\n'
        )

        c_src21 = (
            'int a = 100;' + '\n' +
            'int b = 3;' + '\n' +
            'int mod = 1000000007;' + '\n' +
            'int c = ((a % mod + b % mod) % mod *' \
            '(a % mod - b % mod) % mod) % mod;' + '\n'
        )

        c_src22 = (
            'bool a = 1 == 2, b = 1 != 2;'
        )

        c_src23 = (
            'bool a = 1 < 2, b = 1 <= 2, c = 1 > 2, d = 1 >= 2;'
        )

        c_src24 = (
            'int a = 1, b = 2;' + '\n' +

            'bool c1 = a == 1;' + '\n' +
            'bool c2 = b == 2;' + '\n' +

            'bool c3 = 1 != a;' + '\n' +
            'bool c4 = 1 != b;' + '\n' +

            'bool c5 = a < 0;' + '\n' +
            'bool c6 = b <= 10;' + '\n' +
            'bool c7 = a > 0;' + '\n' +
            'bool c8 = b >= 11;'

        )

        c_src25 = (
            'int a = 3, b = 4;' + '\n' +

            'bool c1 = a == b;' + '\n' +
            'bool c2 = a != b;' + '\n' +
            'bool c3 = a < b;' + '\n' +
            'bool c4 = a <= b;' + '\n' +
            'bool c5 = a > b;' + '\n' +
            'bool c6 = a >= b;'
        )

        c_src26 = (
            'float a = 1.25, b = 2.5;' + '\n' +

            'bool c1 = a == 1.25;' + '\n' +
            'bool c2 = b == 2.54;' + '\n' +

            'bool c3 = 1.2 != a;' + '\n' +
            'bool c4 = 1.5 != b;'
        )

        c_src27 = (
            'float a = 1.25, b = 2.5;' + '\n' +

            'bool c1 = a == b;' + '\n' +
            'bool c2 = a != b;' + '\n' +
            'bool c3 = a < b;' + '\n' +
            'bool c4 = a <= b;' + '\n' +
            'bool c5 = a > b;' + '\n' +
            'bool c6 = a >= b;'
        )

        c_src28 = (
            'bool c1 = true == true;' + '\n' +
            'bool c2 = true == false;' + '\n' +
            'bool c3 = false == false;' + '\n' +

            'bool c4 = true != true;' + '\n' +
            'bool c5 = true != false;' + '\n' +
            'bool c6 = false != false;'
        )

        c_src29 = (
            'bool c1 = true && true;' + '\n' +
            'bool c2 = true && false;' + '\n' +
            'bool c3 = false && false;' + '\n' +

            'bool c4 = true || true;' + '\n' +
            'bool c5 = true || false;' + '\n' +
            'bool c6 = false || false;'
        )

        c_src30 = (
            'bool a = false;' + '\n' +

            'bool c1 = a && true;' + '\n' +
            'bool c2 = false && a;' + '\n' +

            'bool c3 = true || a;' + '\n' +
            'bool c4 = a || false;'
        )

        c_src31 = (
            'int a = 1;' + '\n' +

            'bool c1 = a && 1;' + '\n' +
            'bool c2 = a && 0;' + '\n' +

            'bool c3 = a || 1;' + '\n' +
            'bool c4 = 0 || a;'
        )

        c_src32 = (
            'int a = 1, b = 0;' + '\n' +
            'bool c = false, d = true;'+ '\n' +

            'bool c1 = a && b;' + '\n' +
            'bool c2 = a && c;' + '\n' +
            'bool c3 = c && d;' + '\n' +

            'bool c4 = a || b;' + '\n' +
            'bool c5 = a || c;' + '\n' +
            'bool c6 = c || d;'
        )

        c_src_raise1 = (
            "char a = 'b';"
        )

        c_src_raise2 = (
            'int a[] = {10, 20};'
        )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()
        res3 = SymPyExpression(c_src3, 'c').return_expr()
        res4 = SymPyExpression(c_src4, 'c').return_expr()
        res5 = SymPyExpression(c_src5, 'c').return_expr()
        res6 = SymPyExpression(c_src6, 'c').return_expr()
        res7 = SymPyExpression(c_src7, 'c').return_expr()
        res8 = SymPyExpression(c_src8, 'c').return_expr()
        res9 = SymPyExpression(c_src9, 'c').return_expr()
        res10 = SymPyExpression(c_src10, 'c').return_expr()
        res11 = SymPyExpression(c_src11, 'c').return_expr()
        res12 = SymPyExpression(c_src12, 'c').return_expr()
        res13 = SymPyExpression(c_src13, 'c').return_expr()
        res14 = SymPyExpression(c_src14, 'c').return_expr()
        res15 = SymPyExpression(c_src15, 'c').return_expr()
        res16 = SymPyExpression(c_src16, 'c').return_expr()
        res17 = SymPyExpression(c_src17, 'c').return_expr()
        res18 = SymPyExpression(c_src18, 'c').return_expr()
        res19 = SymPyExpression(c_src19, 'c').return_expr()
        res20 = SymPyExpression(c_src20, 'c').return_expr()
        res21 = SymPyExpression(c_src21, 'c').return_expr()
        res22 = SymPyExpression(c_src22, 'c').return_expr()
        res23 = SymPyExpression(c_src23, 'c').return_expr()
        res24 = SymPyExpression(c_src24, 'c').return_expr()
        res25 = SymPyExpression(c_src25, 'c').return_expr()
        res26 = SymPyExpression(c_src26, 'c').return_expr()
        res27 = SymPyExpression(c_src27, 'c').return_expr()
        res28 = SymPyExpression(c_src28, 'c').return_expr()
        res29 = SymPyExpression(c_src29, 'c').return_expr()
        res30 = SymPyExpression(c_src30, 'c').return_expr()
        res31 = SymPyExpression(c_src31, 'c').return_expr()
        res32 = SymPyExpression(c_src32, 'c').return_expr()

        assert res1[0] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(100)
                )
            )

        assert res1[1] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Symbol('b')
                )
            )

        assert res2[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res2[1] == Declaration(Variable(Symbol('b'),
            type=IntBaseType(String('intc')),
            value=Add(
                Symbol('a'),
                Integer(1)
                )
            )
        )

        assert res3[0] == Declaration(
            Variable(Symbol('a'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('12.5', precision=53)
                )
            )

        assert res3[1] == Declaration(
            Variable(Symbol('b'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Mul(
                    Float('20.0', precision=53),
                    Symbol('a')
                    )
                )
            )

        assert res4[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(83)
                )
            )

        assert res5[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(-4836)
                )
            )

        assert res6[0] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(2)
                )
            )

        assert res6[1] == Declaration(
            Variable(Symbol('c'),
                type=IntBaseType(String('intc')),
                value=Integer(3)
                )
            )

        assert res6[2] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Add(
                    Symbol('b'),
                    Mul(
                        Integer(4),
                        Symbol('c')
                        )
                    )
                )
            )

        assert res7[0] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res7[1] == Declaration(
            Variable(Symbol('c'),
                type=IntBaseType(String('intc')),
                value=Add(
                    Symbol('b'),
                    Integer(2)
                    )
                )
            )

        assert res7[2] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Mul(
                    Integer(10),
                    Pow(
                        Symbol('b'),
                        Integer(2)
                        ),
                    Symbol('c')
                    )
                )
            )

        assert res8[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(1)
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc')),
                        value=Integer(2)
                        )
                    ),
                Declaration(
                    Variable(Symbol('temp'),
                        type=IntBaseType(String('intc')),
                        value=Symbol('a')
                        )
                    ),
                Assignment(
                    Variable(Symbol('a')),
                    Symbol('b')
                    ),
                Assignment(
                    Variable(Symbol('b')),
                    Symbol('temp')
                    )
                )
            )

        assert res9[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res9[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(2)
                )
            )

        assert res9[2] == Declaration(
            Variable(Symbol('c'),
                type=IntBaseType(String('intc')),
                value=Symbol('a')
                )
            )

        assert res9[3] == Declaration(
            Variable(Symbol('d'),
                type=IntBaseType(String('intc')),
                value=Add(
                    Symbol('a'),
                    Symbol('b'),
                    Symbol('c')
                    )
                )
            )

        assert res9[4] == Declaration(
            Variable(Symbol('e'),
                type=IntBaseType(String('intc')),
                value=Add(
                    Pow(
                        Symbol('a'),
                        Integer(3)
                        ),
                    Mul(
                        Integer(3),
                        Pow(
                            Symbol('a'),
                            Integer(2)
                            ),
                        Symbol('b')
                        ),
                    Mul(
                        Integer(3),
                        Symbol('a'),
                        Pow(
                            Symbol('b'),
                            Integer(2)
                            )
                        ),
                    Pow(
                        Symbol('b'),
                        Integer(3)
                        )
                    )
                )
            )

        assert res9[5] == Declaration(
            Variable(Symbol('f'),
                type=IntBaseType(String('intc')),
                value=Mul(
                    Add(
                        Symbol('a'),
                        Symbol('b'),
                        Mul(
                            Integer(-1),
                            Symbol('c')
                            )
                        ),
                    Add(
                        Symbol('a'),
                        Symbol('b'),
                        Symbol('c')
                        )
                    )
                )
            )

        assert res9[6] == Declaration(
            Variable(Symbol('g'),
                type=IntBaseType(String('intc')),
                value=Mul(
                    Symbol('a'),
                    Add(
                        Symbol('b'),
                        Mul(
                            Integer(-1),
                            Symbol('c')
                            )
                        ),
                    Pow(
                        Add(
                            Symbol('a'),
                            Symbol('b'),
                            Symbol('c'),
                            Symbol('d')
                            ),
                        Integer(2)
                        )
                    )
                )
            )

        assert res10[0] == Declaration(
            Variable(Symbol('a'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('10.0', precision=53)
                )
            )

        assert res10[1] == Declaration(
            Variable(Symbol('b'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('2.5', precision=53)
                )
            )

        assert res10[2] == Declaration(
            Variable(Symbol('c'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Add(
                    Pow(
                        Symbol('a'),
                        Integer(2)
                        ),
                    Mul(
                        Integer(2),
                        Symbol('a'),
                        Symbol('b')
                        ),
                    Pow(
                        Symbol('b'),
                        Integer(2)
                        )
                    )
                )
            )

        assert res11[0] == Declaration(
            Variable(Symbol('a'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('4.0', precision=53)
                )
            )

        assert res12[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(25)
                )
            )

        assert res13[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(-95)
                )
            )

        assert res14[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(-300)
                )
            )

        assert res15[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(4)
                )
            )

        assert res15[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(2)
                )
            )

        assert res15[2] == Declaration(
            Variable(Symbol('c'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Mul(
                    Pow(
                        Symbol('a'),
                        Integer(-1)
                        ),
                    Symbol('b')
                    )
                )
            )

        assert res16[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(2)
                )
            )

        assert res16[1] == Declaration(
            Variable(Symbol('d'),
                type=IntBaseType(String('intc')),
                value=Integer(5)
                )
            )

        assert res16[2] == Declaration(
            Variable(Symbol('n'),
                type=IntBaseType(String('intc')),
                value=Integer(10)
                )
            )

        assert res16[3] == Declaration(
            Variable(Symbol('s'),
                type=IntBaseType(String('intc')),
                value=Mul(
                    Rational(1, 2),
                    Symbol('a'),
                    Add(
                        Mul(
                            Integer(2),
                            Symbol('a')
                            ),
                        Mul(
                            Symbol('d'),
                            Add(
                                Symbol('n'),
                                Integer(-1)
                                )
                            )
                        )
                    )
                )
            )

        assert res17[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res18[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(2)
                )
            )

        assert res18[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Mod(
                    Symbol('a'),
                    Integer(3)
                    )
                )
            )

        assert res19[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(100)
                )
            )
        assert res19[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(3)
                )
            )

        assert res19[2] == Declaration(
            Variable(Symbol('c'),
                type=IntBaseType(String('intc')),
                value=Mod(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res20[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(100)
                )
            )

        assert res20[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(3)
                )
            )

        assert res20[2] == Declaration(
            Variable(Symbol('mod'),
                type=IntBaseType(String('intc')),
                value=Integer(1000000007)
                )
            )

        assert res20[3] == Declaration(
            Variable(Symbol('c'),
                type=IntBaseType(String('intc')),
                value=Mod(
                    Add(
                        Symbol('a'),
                        Mul(
                            Integer(100),
                            Pow(
                                Symbol('a'),
                                Integer(-1)
                                ),
                            Symbol('b')
                            )
                        ),
                    Symbol('mod')
                    )
                )
            )

        assert res21[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(100)
                )
            )

        assert res21[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(3)
                )
            )

        assert res21[2] == Declaration(
            Variable(Symbol('mod'),
                type=IntBaseType(String('intc')),
                value=Integer(1000000007)
                )
            )

        assert res21[3] == Declaration(
            Variable(Symbol('c'),
                type=IntBaseType(String('intc')),
                value=Mod(
                    Mul(
                        Add(
                            Symbol('a'),
                            Mul(
                                Integer(-1),
                                Symbol('b')
                                )
                            ),
                        Add(
                            Symbol('a'),
                            Symbol('b')
                            )
                        ),
                    Symbol('mod')
                    )
                )
            )

        assert res22[0] == Declaration(
            Variable(Symbol('a'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res22[1] == Declaration(
            Variable(Symbol('b'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res23[0] == Declaration(
            Variable(Symbol('a'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res23[1] == Declaration(
            Variable(Symbol('b'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res23[2] == Declaration(
            Variable(Symbol('c'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res23[3] == Declaration(
            Variable(Symbol('d'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res24[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res24[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(2)
                )
            )

        assert res24[2] == Declaration(
            Variable(Symbol('c1'),
                type=Type(String('bool')),
                value=Equality(
                    Symbol('a'),
                    Integer(1)
                    )
                )
            )

        assert res24[3] == Declaration(
            Variable(Symbol('c2'),
                type=Type(String('bool')),
                value=Equality(
                    Symbol('b'),
                    Integer(2)
                    )
                )
            )

        assert res24[4] == Declaration(
            Variable(Symbol('c3'),
                type=Type(String('bool')),
                value=Unequality(
                    Integer(1),
                    Symbol('a')
                    )
                )
            )

        assert res24[5] == Declaration(
            Variable(Symbol('c4'),
                type=Type(String('bool')),
                value=Unequality(
                    Integer(1),
                    Symbol('b')
                    )
                )
            )

        assert res24[6] == Declaration(
            Variable(Symbol('c5'),
                type=Type(String('bool')),
                value=StrictLessThan(Symbol('a'),
                    Integer(0)
                    )
                )
            )

        assert res24[7] == Declaration(
            Variable(Symbol('c6'),
                type=Type(String('bool')),
                value=LessThan(
                    Symbol('b'),
                    Integer(10)
                    )
                )
            )

        assert res24[8] == Declaration(
            Variable(Symbol('c7'),
                type=Type(String('bool')),
                value=StrictGreaterThan(
                    Symbol('a'),
                    Integer(0)
                    )
                )
            )

        assert res24[9] == Declaration(
            Variable(Symbol('c8'),
                type=Type(String('bool')),
                value=GreaterThan(
                    Symbol('b'),
                    Integer(11)
                    )
                )
            )

        assert res25[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(3)
                )
            )

        assert res25[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(4)
                )
            )

        assert res25[2] == Declaration(Variable(Symbol('c1'),
            type=Type(String('bool')),
            value=Equality(
                Symbol('a'),
                Symbol('b')
                )
            )
        )

        assert res25[3] == Declaration(
            Variable(Symbol('c2'),
                type=Type(String('bool')),
                value=Unequality(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res25[4] == Declaration(
            Variable(Symbol('c3'),
                type=Type(String('bool')),
                value=StrictLessThan(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res25[5] == Declaration(
            Variable(Symbol('c4'),
                type=Type(String('bool')),
                value=LessThan(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res25[6] == Declaration(
            Variable(Symbol('c5'),
                type=Type(String('bool')),
                value=StrictGreaterThan(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res25[7] == Declaration(
            Variable(Symbol('c6'),
                type=Type(String('bool')),
                value=GreaterThan(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res26[0] == Declaration(
            Variable(Symbol('a'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('1.25', precision=53)
                )
            )

        assert res26[1] == Declaration(
            Variable(Symbol('b'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('2.5', precision=53)
                )
            )

        assert res26[2] == Declaration(
            Variable(Symbol('c1'),
                type=Type(String('bool')),
                value=Equality(
                    Symbol('a'),
                    Float('1.25', precision=53)
                    )
                )
            )

        assert res26[3] == Declaration(
            Variable(Symbol('c2'),
                type=Type(String('bool')),
                value=Equality(
                    Symbol('b'),
                    Float('2.54', precision=53)
                    )
                )
            )

        assert res26[4] == Declaration(
            Variable(Symbol('c3'),
                type=Type(String('bool')),
                value=Unequality(
                    Float('1.2', precision=53),
                    Symbol('a')
                    )
                )
            )

        assert res26[5] == Declaration(
            Variable(Symbol('c4'),
                type=Type(String('bool')),
                value=Unequality(
                    Float('1.5', precision=53),
                    Symbol('b')
                    )
                )
            )

        assert res27[0] == Declaration(
            Variable(Symbol('a'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('1.25', precision=53)
                )
            )

        assert res27[1] == Declaration(
            Variable(Symbol('b'),
                type=FloatType(
                    String('float32'),
                    nbits=Integer(32),
                    nmant=Integer(23),
                    nexp=Integer(8)
                    ),
                value=Float('2.5', precision=53)
                )
            )

        assert res27[2] == Declaration(
            Variable(Symbol('c1'),
                type=Type(String('bool')),
                value=Equality(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res27[3] == Declaration(
            Variable(Symbol('c2'),
                type=Type(String('bool')),
                value=Unequality(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res27[4] == Declaration(
            Variable(Symbol('c3'),
                type=Type(String('bool')),
                value=StrictLessThan(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res27[5] == Declaration(
            Variable(Symbol('c4'),
                type=Type(String('bool')),
                value=LessThan(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res27[6] == Declaration(
            Variable(Symbol('c5'),
                type=Type(String('bool')),
                value=StrictGreaterThan(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res27[7] == Declaration(
            Variable(Symbol('c6'),
                type=Type(String('bool')),
                value=GreaterThan(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res28[0] == Declaration(
            Variable(Symbol('c1'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res28[1] == Declaration(
            Variable(Symbol('c2'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res28[2] == Declaration(
            Variable(Symbol('c3'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res28[3] == Declaration(
            Variable(Symbol('c4'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res28[4] == Declaration(
            Variable(Symbol('c5'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res28[5] == Declaration(
            Variable(Symbol('c6'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res29[0] == Declaration(
            Variable(Symbol('c1'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res29[1] == Declaration(
            Variable(Symbol('c2'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res29[2] == Declaration(
            Variable(Symbol('c3'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res29[3] == Declaration(
            Variable(Symbol('c4'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res29[4] == Declaration(
            Variable(Symbol('c5'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res29[5] == Declaration(
            Variable(Symbol('c6'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res30[0] == Declaration(
            Variable(Symbol('a'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res30[1] == Declaration(
            Variable(Symbol('c1'),
                type=Type(String('bool')),
                value=Symbol('a')
                )
            )

        assert res30[2] == Declaration(
            Variable(Symbol('c2'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res30[3] == Declaration(
            Variable(Symbol('c3'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res30[4] == Declaration(
            Variable(Symbol('c4'),
                type=Type(String('bool')),
                value=Symbol('a')
                )
            )

        assert res31[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res31[1] == Declaration(
            Variable(Symbol('c1'),
                type=Type(String('bool')),
                value=Symbol('a')
                )
            )

        assert res31[2] == Declaration(
            Variable(Symbol('c2'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res31[3] == Declaration(
            Variable(Symbol('c3'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res31[4] == Declaration(
            Variable(Symbol('c4'),
                type=Type(String('bool')),
                value=Symbol('a')
                )
            )

        assert res32[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res32[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(0)
                )
            )

        assert res32[2] == Declaration(
            Variable(Symbol('c'),
                type=Type(String('bool')),
                value=false
                )
            )

        assert res32[3] == Declaration(
            Variable(Symbol('d'),
                type=Type(String('bool')),
                value=true
                )
            )

        assert res32[4] == Declaration(
            Variable(Symbol('c1'),
                type=Type(String('bool')),
                value=And(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res32[5] == Declaration(
            Variable(Symbol('c2'),
                type=Type(String('bool')),
                value=And(
                    Symbol('a'),
                    Symbol('c')
                    )
                )
            )

        assert res32[6] == Declaration(
            Variable(Symbol('c3'),
                type=Type(String('bool')),
                value=And(
                    Symbol('c'),
                    Symbol('d')
                    )
                )
            )

        assert res32[7] == Declaration(
            Variable(Symbol('c4'),
                type=Type(String('bool')),
                value=Or(
                    Symbol('a'),
                    Symbol('b')
                    )
                )
            )

        assert res32[8] == Declaration(
            Variable(Symbol('c5'),
                type=Type(String('bool')),
                value=Or(
                    Symbol('a'),
                    Symbol('c')
                    )
                )
            )

        assert res32[9] == Declaration(
            Variable(Symbol('c6'),
                type=Type(String('bool')),
                value=Or(
                    Symbol('c'),
                    Symbol('d')
                    )
                )
            )

        raises(NotImplementedError, lambda: SymPyExpression(c_src_raise1, 'c'))
        raises(NotImplementedError, lambda: SymPyExpression(c_src_raise2, 'c'))


    def test_paren_expr():
        c_src1 = (
            'int a = (1);'
            'int b = (1 + 2 * 3);'
        )

        c_src2 = (
            'int a = 1, b = 2, c = 3;'
            'int d = (a);'
            'int e = (a + 1);'
            'int f = (a + b * c - d / e);'
        )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()

        assert res1[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res1[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(7)
                )
            )

        assert res2[0] == Declaration(
            Variable(Symbol('a'),
                type=IntBaseType(String('intc')),
                value=Integer(1)
                )
            )

        assert res2[1] == Declaration(
            Variable(Symbol('b'),
                type=IntBaseType(String('intc')),
                value=Integer(2)
                )
            )

        assert res2[2] == Declaration(
            Variable(Symbol('c'),
                type=IntBaseType(String('intc')),
                value=Integer(3)
                )
            )

        assert res2[3] == Declaration(
            Variable(Symbol('d'),
                type=IntBaseType(String('intc')),
                value=Symbol('a')
                )
            )

        assert res2[4] == Declaration(
            Variable(Symbol('e'),
                type=IntBaseType(String('intc')),
                value=Add(
                    Symbol('a'),
                    Integer(1)
                    )
                )
            )

        assert res2[5] == Declaration(
            Variable(Symbol('f'),
                type=IntBaseType(String('intc')),
                value=Add(
                    Symbol('a'),
                    Mul(
                        Symbol('b'),
                        Symbol('c')
                        ),
                    Mul(
                        Integer(-1),
                        Symbol('d'),
                        Pow(
                            Symbol('e'),
                            Integer(-1)
                            )
                        )
                    )
                )
            )


    def test_unary_operators():
        c_src1 = (
            'void func()'+
            '{' + '\n' +
                'int a = 10;' + '\n' +
                'int b = 20;' + '\n' +
                '++a;' + '\n' +
                '--b;' + '\n' +
                'a++;' + '\n' +
                'b--;' + '\n' +
            '}'
        )

        c_src2 = (
            'void func()'+
            '{' + '\n' +
                'int a = 10;' + '\n' +
                'int b = -100;' + '\n' +
                'int c = +19;' + '\n' +
                'int d = ++a;' + '\n' +
                'int e = --b;' + '\n' +
                'int f = a++;' + '\n' +
                'int g = b--;' + '\n' +
                'bool h = !false;' + '\n' +
                'bool i = !d;' + '\n' +
                'bool j = !0;' + '\n' +
                'bool k = !10.0;' + '\n' +
            '}'
        )

        c_src_raise1 = (
            'void func()'+
            '{' + '\n' +
                'int a = 10;' + '\n' +
                'int b = ~a;' + '\n' +
            '}'
        )

        c_src_raise2 = (
            'void func()'+
            '{' + '\n' +
                'int a = 10;' + '\n' +
                'int b = *&a;' + '\n' +
            '}'
        )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()

        assert res1[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(10)
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc')),
                        value=Integer(20)
                        )
                    ),
                PreIncrement(Symbol('a')),
                PreDecrement(Symbol('b')),
                PostIncrement(Symbol('a')),
                PostDecrement(Symbol('b'))
                )
            )

        assert res2[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(10)
                        )
                    ),
                Declaration(
                    Variable(Symbol('b'),
                        type=IntBaseType(String('intc')),
                        value=Integer(-100)
                        )
                    ),
                Declaration(
                    Variable(Symbol('c'),
                        type=IntBaseType(String('intc')),
                        value=Integer(19)
                        )
                    ),
                Declaration(
                    Variable(Symbol('d'),
                        type=IntBaseType(String('intc')),
                        value=PreIncrement(Symbol('a'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('e'),
                        type=IntBaseType(String('intc')),
                        value=PreDecrement(Symbol('b'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('f'),
                        type=IntBaseType(String('intc')),
                        value=PostIncrement(Symbol('a'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('g'),
                        type=IntBaseType(String('intc')),
                        value=PostDecrement(Symbol('b'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('h'),
                        type=Type(String('bool')),
                        value=true
                        )
                    ),
                Declaration(
                    Variable(Symbol('i'),
                        type=Type(String('bool')),
                        value=Not(Symbol('d'))
                        )
                    ),
                Declaration(
                    Variable(Symbol('j'),
                        type=Type(String('bool')),
                        value=true
                        )
                    ),
                Declaration(
                    Variable(Symbol('k'),
                        type=Type(String('bool')),
                        value=false
                        )
                    )
                )
            )

        raises(NotImplementedError, lambda: SymPyExpression(c_src_raise1, 'c'))
        raises(NotImplementedError, lambda: SymPyExpression(c_src_raise2, 'c'))


    def test_compound_assignment_operator():
        c_src = (
            'void func()'+
            '{' + '\n' +
                'int a = 100;' + '\n' +
                'a += 10;' + '\n' +
                'a -= 10;' + '\n' +
                'a *= 10;' + '\n' +
                'a /= 10;' + '\n' +
                'a %= 10;' + '\n' +
            '}'
        )

        res = SymPyExpression(c_src, 'c').return_expr()

        assert res[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(
                        Symbol('a'),
                        type=IntBaseType(String('intc')),
                        value=Integer(100)
                        )
                    ),
                AddAugmentedAssignment(
                    Variable(Symbol('a')),
                    Integer(10)
                    ),
                SubAugmentedAssignment(
                    Variable(Symbol('a')),
                    Integer(10)
                    ),
                MulAugmentedAssignment(
                    Variable(Symbol('a')),
                    Integer(10)
                    ),
                DivAugmentedAssignment(
                    Variable(Symbol('a')),
                    Integer(10)
                    ),
                ModAugmentedAssignment(
                    Variable(Symbol('a')),
                    Integer(10)
                    )
                )
            )

    @XFAIL # this is expected to fail because of a bug in the C parser.
    def test_while_stmt():
        c_src1 = (
            'void func()'+
            '{' + '\n' +
                'int i = 0;' + '\n' +
                'while(i < 10)' + '\n' +
                '{' + '\n' +
                    'i++;' + '\n' +
                '}'
            '}'
        )

        c_src2 = (
            'void func()'+
            '{' + '\n' +
                'int i = 0;' + '\n' +
                'while(i < 10)' + '\n' +
                    'i++;' + '\n' +
            '}'
        )

        c_src3 = (
            'void func()'+
            '{' + '\n' +
                'int i = 10;' + '\n' +
                'int cnt = 0;' + '\n' +
                'while(i > 0)' + '\n' +
                '{' + '\n' +
                    'i--;' + '\n' +
                    'cnt++;' + '\n' +
                '}' + '\n' +
            '}'
        )

        c_src4 = (
            'int digit_sum(int n)'+
            '{' + '\n' +
                'int sum = 0;' + '\n' +
                'while(n > 0)' + '\n' +
                '{' + '\n' +
                    'sum += (n % 10);' + '\n' +
                    'n /= 10;' + '\n' +
                '}' + '\n' +
                'return sum;' + '\n' +
            '}'
        )

        c_src5 = (
            'void func()'+
            '{' + '\n' +
                'while(1);' + '\n' +
            '}'
        )

        res1 = SymPyExpression(c_src1, 'c').return_expr()
        res2 = SymPyExpression(c_src2, 'c').return_expr()
        res3 = SymPyExpression(c_src3, 'c').return_expr()
        res4 = SymPyExpression(c_src4, 'c').return_expr()
        res5 = SymPyExpression(c_src5, 'c').return_expr()

        assert res1[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(Symbol('i'),
                        type=IntBaseType(String('intc')),
                        value=Integer(0)
                        )
                    ),
                While(
                    StrictLessThan(
                        Symbol('i'),
                        Integer(10)
                        ),
                    body=CodeBlock(
                        PostIncrement(
                            Symbol('i')
                            )
                        )
                    )
                )
            )

        assert res2[0] == res1[0]

        assert res3[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                Declaration(
                    Variable(
                        Symbol('i'),
                        type=IntBaseType(String('intc')),
                        value=Integer(10)
                        )
                    ),
                Declaration(
                    Variable(
                        Symbol('cnt'),
                        type=IntBaseType(String('intc')),
                        value=Integer(0)
                        )
                    ),
                While(
                    StrictGreaterThan(
                        Symbol('i'),
                        Integer(0)
                        ),
                    body=CodeBlock(
                        PostDecrement(
                            Symbol('i')
                            ),
                        PostIncrement(
                            Symbol('cnt')
                            )
                        )
                    )
                )
            )

        assert res4[0] == FunctionDefinition(
            IntBaseType(String('intc')),
            name=String('digit_sum'),
            parameters=(
                Variable(
                    Symbol('n'),
                    type=IntBaseType(String('intc'))
                    ),
                ),
            body=CodeBlock(
                Declaration(
                    Variable(
                        Symbol('sum'),
                        type=IntBaseType(String('intc')),
                        value=Integer(0)
                        )
                    ),
                While(
                    StrictGreaterThan(
                        Symbol('n'),
                        Integer(0)
                        ),
                    body=CodeBlock(
                        AddAugmentedAssignment(
                            Variable(
                                Symbol('sum')
                                ),
                            Mod(
                                Symbol('n'),
                                Integer(10)
                                )
                            ),
                        DivAugmentedAssignment(
                            Variable(
                                Symbol('n')
                                ),
                            Integer(10)
                            )
                        )
                    ),
                Return('sum')
                )
            )

        assert res5[0] == FunctionDefinition(
            NoneToken(),
            name=String('func'),
            parameters=(),
            body=CodeBlock(
                While(
                    Integer(1),
                    body=CodeBlock(
                        NoneToken()
                        )
                    )
                )
            )


else:
    def test_raise():
        from sympy.parsing.c.c_parser import CCodeConverter
        raises(ImportError, lambda: CCodeConverter())
        raises(ImportError, lambda: SymPyExpression(' ', mode = 'c'))