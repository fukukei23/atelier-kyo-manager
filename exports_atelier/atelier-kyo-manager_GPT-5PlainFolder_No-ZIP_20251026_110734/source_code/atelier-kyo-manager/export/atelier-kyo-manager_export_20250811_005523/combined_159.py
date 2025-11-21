
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\utils\hashes.py ===
import hashlib
from typing import TYPE_CHECKING, BinaryIO, Dict, Iterable, List, NoReturn, Optional

from pip._internal.exceptions import HashMismatch, HashMissing, InstallationError
from pip._internal.utils.misc import read_chunks

if TYPE_CHECKING:
    from hashlib import _Hash


# The recommended hash algo of the moment. Change this whenever the state of
# the art changes; it won't hurt backward compatibility.
FAVORITE_HASH = "sha256"


# Names of hashlib algorithms allowed by the --hash option and ``pip hash``
# Currently, those are the ones at least as collision-resistant as sha256.
STRONG_HASHES = ["sha256", "sha384", "sha512"]


class Hashes:
    """A wrapper that builds multiple hashes at once and checks them against
    known-good values

    """

    def __init__(self, hashes: Optional[Dict[str, List[str]]] = None) -> None:
        """
        :param hashes: A dict of algorithm names pointing to lists of allowed
            hex digests
        """
        allowed = {}
        if hashes is not None:
            for alg, keys in hashes.items():
                # Make sure values are always sorted (to ease equality checks)
                allowed[alg] = [k.lower() for k in sorted(keys)]
        self._allowed = allowed

    def __and__(self, other: "Hashes") -> "Hashes":
        if not isinstance(other, Hashes):
            return NotImplemented

        # If either of the Hashes object is entirely empty (i.e. no hash
        # specified at all), all hashes from the other object are allowed.
        if not other:
            return self
        if not self:
            return other

        # Otherwise only hashes that present in both objects are allowed.
        new = {}
        for alg, values in other._allowed.items():
            if alg not in self._allowed:
                continue
            new[alg] = [v for v in values if v in self._allowed[alg]]
        return Hashes(new)

    @property
    def digest_count(self) -> int:
        return sum(len(digests) for digests in self._allowed.values())

    def is_hash_allowed(self, hash_name: str, hex_digest: str) -> bool:
        """Return whether the given hex digest is allowed."""
        return hex_digest in self._allowed.get(hash_name, [])

    def check_against_chunks(self, chunks: Iterable[bytes]) -> None:
        """Check good hashes against ones built from iterable of chunks of
        data.

        Raise HashMismatch if none match.

        """
        gots = {}
        for hash_name in self._allowed.keys():
            try:
                gots[hash_name] = hashlib.new(hash_name)
            except (ValueError, TypeError):
                raise InstallationError(f"Unknown hash name: {hash_name}")

        for chunk in chunks:
            for hash in gots.values():
                hash.update(chunk)

        for hash_name, got in gots.items():
            if got.hexdigest() in self._allowed[hash_name]:
                return
        self._raise(gots)

    def _raise(self, gots: Dict[str, "_Hash"]) -> "NoReturn":
        raise HashMismatch(self._allowed, gots)

    def check_against_file(self, file: BinaryIO) -> None:
        """Check good hashes against a file-like object

        Raise HashMismatch if none match.

        """
        return self.check_against_chunks(read_chunks(file))

    def check_against_path(self, path: str) -> None:
        with open(path, "rb") as file:
            return self.check_against_file(file)

    def has_one_of(self, hashes: Dict[str, str]) -> bool:
        """Return whether any of the given hashes are allowed."""
        for hash_name, hex_digest in hashes.items():
            if self.is_hash_allowed(hash_name, hex_digest):
                return True
        return False

    def __bool__(self) -> bool:
        """Return whether I know any known-good hashes."""
        return bool(self._allowed)

    def __eq__(self, other: object) -> bool:
        if not isinstance(other, Hashes):
            return NotImplemented
        return self._allowed == other._allowed

    def __hash__(self) -> int:
        return hash(
            ",".join(
                sorted(
                    ":".join((alg, digest))
                    for alg, digest_list in self._allowed.items()
                    for digest in digest_list
                )
            )
        )


class MissingHashes(Hashes):
    """A workalike for Hashes used when we're missing a hash for a requirement

    It computes the actual hash of the requirement and raises a HashMissing
    exception showing it to the user.

    """

    def __init__(self) -> None:
        """Don't offer the ``hashes`` kwarg."""
        # Pass our favorite hash in to generate a "gotten hash". With the
        # empty list, it will never match, so an error will always raise.
        super().__init__(hashes={FAVORITE_HASH: []})

    def _raise(self, gots: Dict[str, "_Hash"]) -> "NoReturn":
        raise HashMissing(gots[FAVORITE_HASH].hexdigest())

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\utils\setuptools_build.py ===
import sys
import textwrap
from typing import List, Optional, Sequence

# Shim to wrap setup.py invocation with setuptools
# Note that __file__ is handled via two {!r} *and* %r, to ensure that paths on
# Windows are correctly handled (it should be "C:\\Users" not "C:\Users").
_SETUPTOOLS_SHIM = textwrap.dedent(
    """
    exec(compile('''
    # This is <pip-setuptools-caller> -- a caller that pip uses to run setup.py
    #
    # - It imports setuptools before invoking setup.py, to enable projects that directly
    #   import from `distutils.core` to work with newer packaging standards.
    # - It provides a clear error message when setuptools is not installed.
    # - It sets `sys.argv[0]` to the underlying `setup.py`, when invoking `setup.py` so
    #   setuptools doesn't think the script is `-c`. This avoids the following warning:
    #     manifest_maker: standard file '-c' not found".
    # - It generates a shim setup.py, for handling setup.cfg-only projects.
    import os, sys, tokenize, traceback

    try:
        import setuptools
    except ImportError:
        print(
            "ERROR: Can not execute `setup.py` since setuptools failed to import in "
            "the build environment with exception:",
            file=sys.stderr,
        )
        traceback.print_exc()
        sys.exit(1)

    __file__ = %r
    sys.argv[0] = __file__

    if os.path.exists(__file__):
        filename = __file__
        with tokenize.open(__file__) as f:
            setup_py_code = f.read()
    else:
        filename = "<auto-generated setuptools caller>"
        setup_py_code = "from setuptools import setup; setup()"

    exec(compile(setup_py_code, filename, "exec"))
    ''' % ({!r},), "<pip-setuptools-caller>", "exec"))
    """
).rstrip()


def make_setuptools_shim_args(
    setup_py_path: str,
    global_options: Optional[Sequence[str]] = None,
    no_user_config: bool = False,
    unbuffered_output: bool = False,
) -> List[str]:
    """
    Get setuptools command arguments with shim wrapped setup file invocation.

    :param setup_py_path: The path to setup.py to be wrapped.
    :param global_options: Additional global options.
    :param no_user_config: If True, disables personal user configuration.
    :param unbuffered_output: If True, adds the unbuffered switch to the
     argument list.
    """
    args = [sys.executable]
    if unbuffered_output:
        args += ["-u"]
    args += ["-c", _SETUPTOOLS_SHIM.format(setup_py_path)]
    if global_options:
        args += global_options
    if no_user_config:
        args += ["--no-user-cfg"]
    return args


def make_setuptools_bdist_wheel_args(
    setup_py_path: str,
    global_options: Sequence[str],
    build_options: Sequence[str],
    destination_dir: str,
) -> List[str]:
    # NOTE: Eventually, we'd want to also -S to the flags here, when we're
    # isolating. Currently, it breaks Python in virtualenvs, because it
    # relies on site.py to find parts of the standard library outside the
    # virtualenv.
    args = make_setuptools_shim_args(
        setup_py_path, global_options=global_options, unbuffered_output=True
    )
    args += ["bdist_wheel", "-d", destination_dir]
    args += build_options
    return args


def make_setuptools_clean_args(
    setup_py_path: str,
    global_options: Sequence[str],
) -> List[str]:
    args = make_setuptools_shim_args(
        setup_py_path, global_options=global_options, unbuffered_output=True
    )
    args += ["clean", "--all"]
    return args


def make_setuptools_develop_args(
    setup_py_path: str,
    *,
    global_options: Sequence[str],
    no_user_config: bool,
    prefix: Optional[str],
    home: Optional[str],
    use_user_site: bool,
) -> List[str]:
    assert not (use_user_site and prefix)

    args = make_setuptools_shim_args(
        setup_py_path,
        global_options=global_options,
        no_user_config=no_user_config,
    )

    args += ["develop", "--no-deps"]

    if prefix:
        args += ["--prefix", prefix]
    if home is not None:
        args += ["--install-dir", home]

    if use_user_site:
        args += ["--user", "--prefix="]

    return args


def make_setuptools_egg_info_args(
    setup_py_path: str,
    egg_info_dir: Optional[str],
    no_user_config: bool,
) -> List[str]:
    args = make_setuptools_shim_args(setup_py_path, no_user_config=no_user_config)

    args += ["egg_info"]

    if egg_info_dir:
        args += ["--egg-base", egg_info_dir]

    return args

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\sets\handlers\union.py ===
from sympy.core.singleton import S
from sympy.core.sympify import sympify
from sympy.functions.elementary.miscellaneous import Min, Max
from sympy.sets.sets import (EmptySet, FiniteSet, Intersection,
    Interval, ProductSet, Set, Union, UniversalSet)
from sympy.sets.fancysets import (ComplexRegion, Naturals, Naturals0,
    Integers, Rationals, Reals)
from sympy.multipledispatch import Dispatcher


union_sets = Dispatcher('union_sets')


@union_sets.register(Naturals0, Naturals)
def _(a, b):
    return a

@union_sets.register(Rationals, Naturals)
def _(a, b):
    return a

@union_sets.register(Rationals, Naturals0)
def _(a, b):
    return a

@union_sets.register(Reals, Naturals)
def _(a, b):
    return a

@union_sets.register(Reals, Naturals0)
def _(a, b):
    return a

@union_sets.register(Reals, Rationals)
def _(a, b):
    return a

@union_sets.register(Integers, Set)
def _(a, b):
    intersect = Intersection(a, b)
    if intersect == a:
        return b
    elif intersect == b:
        return a

@union_sets.register(ComplexRegion, Set)
def _(a, b):
    if b.is_subset(S.Reals):
        # treat a subset of reals as a complex region
        b = ComplexRegion.from_real(b)

    if b.is_ComplexRegion:
        # a in rectangular form
        if (not a.polar) and (not b.polar):
            return ComplexRegion(Union(a.sets, b.sets))
        # a in polar form
        elif a.polar and b.polar:
            return ComplexRegion(Union(a.sets, b.sets), polar=True)
    return None

@union_sets.register(EmptySet, Set)
def _(a, b):
    return b


@union_sets.register(UniversalSet, Set)
def _(a, b):
    return a

@union_sets.register(ProductSet, ProductSet)
def _(a, b):
    if b.is_subset(a):
        return a
    if len(b.sets) != len(a.sets):
        return None
    if len(a.sets) == 2:
        a1, a2 = a.sets
        b1, b2 = b.sets
        if a1 == b1:
            return a1 * Union(a2, b2)
        if a2 == b2:
            return Union(a1, b1) * a2
    return None

@union_sets.register(ProductSet, Set)
def _(a, b):
    if b.is_subset(a):
        return a
    return None

@union_sets.register(Interval, Interval)
def _(a, b):
    if a._is_comparable(b):
        # Non-overlapping intervals
        end = Min(a.end, b.end)
        start = Max(a.start, b.start)
        if (end < start or
           (end == start and (end not in a and end not in b))):
            return None
        else:
            start = Min(a.start, b.start)
            end = Max(a.end, b.end)

            left_open = ((a.start != start or a.left_open) and
                         (b.start != start or b.left_open))
            right_open = ((a.end != end or a.right_open) and
                          (b.end != end or b.right_open))
            return Interval(start, end, left_open, right_open)

@union_sets.register(Interval, UniversalSet)
def _(a, b):
    return S.UniversalSet

@union_sets.register(Interval, Set)
def _(a, b):
    # If I have open end points and these endpoints are contained in b
    # But only in case, when endpoints are finite. Because
    # interval does not contain oo or -oo.
    open_left_in_b_and_finite = (a.left_open and
                                     sympify(b.contains(a.start)) is S.true and
                                     a.start.is_finite)
    open_right_in_b_and_finite = (a.right_open and
                                      sympify(b.contains(a.end)) is S.true and
                                      a.end.is_finite)
    if open_left_in_b_and_finite or open_right_in_b_and_finite:
        # Fill in my end points and return
        open_left = a.left_open and a.start not in b
        open_right = a.right_open and a.end not in b
        new_a = Interval(a.start, a.end, open_left, open_right)
        return {new_a, b}
    return None

@union_sets.register(FiniteSet, FiniteSet)
def _(a, b):
    return FiniteSet(*(a._elements | b._elements))

@union_sets.register(FiniteSet, Set)
def _(a, b):
    # If `b` set contains one of my elements, remove it from `a`
    if any(b.contains(x) == True for x in a):
        return {
            FiniteSet(*[x for x in a if b.contains(x) != True]), b}
    return None

@union_sets.register(Set, Set)
def _(a, b):
    return None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\flask_wtf\file.py ===
from collections import abc

from werkzeug.datastructures import FileStorage
from wtforms import FileField as _FileField
from wtforms import MultipleFileField as _MultipleFileField
from wtforms.validators import DataRequired
from wtforms.validators import StopValidation
from wtforms.validators import ValidationError


class FileField(_FileField):
    """Werkzeug-aware subclass of :class:`wtforms.fields.FileField`."""

    def process_formdata(self, valuelist):
        valuelist = (x for x in valuelist if isinstance(x, FileStorage) and x)
        data = next(valuelist, None)

        if data is not None:
            self.data = data
        else:
            self.raw_data = ()


class MultipleFileField(_MultipleFileField):
    """Werkzeug-aware subclass of :class:`wtforms.fields.MultipleFileField`.

    .. versionadded:: 1.2.0
    """

    def process_formdata(self, valuelist):
        valuelist = (x for x in valuelist if isinstance(x, FileStorage) and x)
        data = list(valuelist) or None

        if data is not None:
            self.data = data
        else:
            self.raw_data = ()


class FileRequired(DataRequired):
    """Validates that the uploaded files(s) is a Werkzeug
    :class:`~werkzeug.datastructures.FileStorage` object.

    :param message: error message

    You can also use the synonym ``file_required``.
    """

    def __call__(self, form, field):
        field_data = [field.data] if not isinstance(field.data, list) else field.data
        if not (
            all(isinstance(x, FileStorage) and x for x in field_data) and field_data
        ):
            raise StopValidation(
                self.message or field.gettext("This field is required.")
            )


file_required = FileRequired


class FileAllowed:
    """Validates that the uploaded file(s) is allowed by a given list of
    extensions or a Flask-Uploads :class:`~flaskext.uploads.UploadSet`.

    :param upload_set: A list of extensions or an
        :class:`~flaskext.uploads.UploadSet`
    :param message: error message

    You can also use the synonym ``file_allowed``.
    """

    def __init__(self, upload_set, message=None):
        self.upload_set = upload_set
        self.message = message

    def __call__(self, form, field):
        field_data = [field.data] if not isinstance(field.data, list) else field.data
        if not (
            all(isinstance(x, FileStorage) and x for x in field_data) and field_data
        ):
            return

        filenames = [f.filename.lower() for f in field_data]

        for filename in filenames:
            if isinstance(self.upload_set, abc.Iterable):
                if any(filename.endswith("." + x) for x in self.upload_set):
                    continue

                raise StopValidation(
                    self.message
                    or field.gettext(
                        "File does not have an approved extension: {extensions}"
                    ).format(extensions=", ".join(self.upload_set))
                )

            if not self.upload_set.file_allowed(field_data, filename):
                raise StopValidation(
                    self.message
                    or field.gettext("File does not have an approved extension.")
                )


file_allowed = FileAllowed


class FileSize:
    """Validates that the uploaded file(s) is within a minimum and maximum
    file size (set in bytes).

    :param min_size: minimum allowed file size (in bytes). Defaults to 0 bytes.
    :param max_size: maximum allowed file size (in bytes).
    :param message: error message

    You can also use the synonym ``file_size``.
    """

    def __init__(self, max_size, min_size=0, message=None):
        self.min_size = min_size
        self.max_size = max_size
        self.message = message

    def __call__(self, form, field):
        field_data = [field.data] if not isinstance(field.data, list) else field.data
        if not (
            all(isinstance(x, FileStorage) and x for x in field_data) and field_data
        ):
            return

        for f in field_data:
            file_size = len(f.read())
            f.seek(0)  # reset cursor position to beginning of file

            if (file_size < self.min_size) or (file_size > self.max_size):
                # the file is too small or too big => validation failure
                raise ValidationError(
                    self.message
                    or field.gettext(
                        f"File must be between {self.min_size}"
                        f" and {self.max_size} bytes."
                    )
                )


file_size = FileSize

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\humanfriendly\compat.py ===
# Human friendly input/output in Python.
#
# Author: Peter Odding <peter@peterodding.com>
# Last Change: September 17, 2021
# URL: https://humanfriendly.readthedocs.io

"""
Compatibility with Python 2 and 3.

This module exposes aliases and functions that make it easier to write Python
code that is compatible with Python 2 and Python 3.

.. data:: basestring

   Alias for :func:`python2:basestring` (in Python 2) or :class:`python3:str`
   (in Python 3). See also :func:`is_string()`.

.. data:: HTMLParser

   Alias for :class:`python2:HTMLParser.HTMLParser` (in Python 2) or
   :class:`python3:html.parser.HTMLParser` (in Python 3).

.. data:: interactive_prompt

   Alias for :func:`python2:raw_input()` (in Python 2) or
   :func:`python3:input()` (in Python 3).

.. data:: StringIO

   Alias for :class:`python2:StringIO.StringIO` (in Python 2) or
   :class:`python3:io.StringIO` (in Python 3).

.. data:: unicode

   Alias for :func:`python2:unicode` (in Python 2) or :class:`python3:str` (in
   Python 3). See also :func:`coerce_string()`.

.. data:: monotonic

   Alias for :func:`python3:time.monotonic()` (in Python 3.3 and higher) or
   `monotonic.monotonic()` (a `conditional dependency
   <https://pypi.org/project/monotonic/>`_ on older Python versions).
"""

__all__ = (
    'HTMLParser',
    'StringIO',
    'basestring',
    'coerce_string',
    'interactive_prompt',
    'is_string',
    'is_unicode',
    'monotonic',
    'name2codepoint',
    'on_macos',
    'on_windows',
    'unichr',
    'unicode',
    'which',
)

# Standard library modules.
import sys

# Differences between Python 2 and 3.
try:
    # Python 2.
    unicode = unicode
    unichr = unichr
    basestring = basestring
    interactive_prompt = raw_input
    from distutils.spawn import find_executable as which
    from HTMLParser import HTMLParser
    from StringIO import StringIO
    from htmlentitydefs import name2codepoint
except (ImportError, NameError):
    # Python 3.
    unicode = str
    unichr = chr
    basestring = str
    interactive_prompt = input
    from shutil import which
    from html.parser import HTMLParser
    from io import StringIO
    from html.entities import name2codepoint

try:
    # Python 3.3 and higher.
    from time import monotonic
except ImportError:
    # A replacement for older Python versions:
    # https://pypi.org/project/monotonic/
    try:
        from monotonic import monotonic
    except (ImportError, RuntimeError):
        # We fall back to the old behavior of using time.time() instead of
        # failing when {time,monotonic}.monotonic() are both missing.
        from time import time as monotonic


def coerce_string(value):
    """
    Coerce any value to a Unicode string (:func:`python2:unicode` in Python 2 and :class:`python3:str` in Python 3).

    :param value: The value to coerce.
    :returns: The value coerced to a Unicode string.
    """
    return value if is_string(value) else unicode(value)


def is_string(value):
    """
    Check if a value is a :func:`python2:basestring` (in Python 2) or :class:`python3:str` (in Python 3) object.

    :param value: The value to check.
    :returns: :data:`True` if the value is a string, :data:`False` otherwise.
    """
    return isinstance(value, basestring)


def is_unicode(value):
    """
    Check if a value is a :func:`python2:unicode` (in Python 2) or :class:`python2:str` (in Python 3) object.

    :param value: The value to check.
    :returns: :data:`True` if the value is a Unicode string, :data:`False` otherwise.
    """
    return isinstance(value, unicode)


def on_macos():
    """
    Check if we're running on Apple MacOS.

    :returns: :data:`True` if running MacOS, :data:`False` otherwise.
    """
    return sys.platform.startswith('darwin')


def on_windows():
    """
    Check if we're running on the Microsoft Windows OS.

    :returns: :data:`True` if running Windows, :data:`False` otherwise.
    """
    return sys.platform.startswith('win')

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\gpt2\parity_check_helper.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.  See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------
# This script helps debugging parity issue for two same onnx models with fp16 and fp32 format
# Please build ORT with --cmake_extra_defines onnxruntime_DEBUG_NODE_INPUTS_OUTPUTS=ON

import math
import multiprocessing
import os
from pathlib import Path

import numpy
import torch
from benchmark_helper import create_onnxruntime_session
from gpt2_helper import Gpt2Helper
from onnx import TensorProto, numpy_helper

NON_ZERO_VALUE = str(1)
ZERO_VALUE = str(0)


def environ_setting_nodes(node_name_filter=None, node_type_filter=None):
    # Set I/O data as default
    os.environ["ORT_DEBUG_NODE_IO_DUMP_SHAPE_DATA"] = ZERO_VALUE
    os.environ["ORT_DEBUG_NODE_IO_DUMP_INPUT_DATA"] = NON_ZERO_VALUE
    os.environ["ORT_DEBUG_NODE_IO_DUMP_OUTPUT_DATA"] = NON_ZERO_VALUE
    if node_name_filter is not None:
        os.environ["ORT_DEBUG_NODE_IO_NAME_FILTER"] = node_name_filter
    elif node_type_filter is not None:
        os.environ["ORT_DEBUG_NODE_IO_OP_TYPE_FILTER"] = node_type_filter
    else:
        os.environ["ORT_DEBUG_NODE_IO_DUMPING_DATA_TO_FILES_FOR_ALL_NODES_IS_OK"] = NON_ZERO_VALUE


def environ_setting_paths(output_path):
    # Set dumping values to files as default
    os.environ["ORT_DEBUG_NODE_IO_DUMP_DATA_DESTINATION"] = "files"
    os.environ["ORT_DEBUG_NODE_IO_OUTPUT_DIR"] = output_path


def environ_reset():
    for flag in [
        "ORT_DEBUG_NODE_IO_DUMP_SHAPE_DATA",
        "ORT_DEBUG_NODE_IO_DUMP_INPUT_DATA",
        "ORT_DEBUG_NODE_IO_DUMP_OUTPUT_DATA",
        "ORT_DEBUG_NODE_IO_NAME_FILTER",
        "ORT_DEBUG_NODE_IO_OP_TYPE_FILTER",
        "ORT_DEBUG_NODE_IO_DUMP_DATA_TO_FILES",
        "ORT_DEBUG_NODE_IO_OUTPUT_DIR",
        "ORT_DEBUG_NODE_IO_DUMPING_DATA_TO_FILES_FOR_ALL_NODES_IS_OK",
    ]:
        if flag in os.environ:
            del os.environ[flag]


def inference(model_path, dummy_inputs, outputs_path, use_gpu):
    environ_reset()
    environ_setting_nodes()
    environ_setting_paths(outputs_path)
    session = create_onnxruntime_session(model_path, use_gpu, enable_all_optimization=False)
    Gpt2Helper.onnxruntime_inference(session, dummy_inputs)


def generate_outputs_files(model_path, dummy_inputs, outputs_path, use_gpu):
    dir_path = Path(outputs_path)
    if dir_path.exists() and dir_path.is_dir():
        import shutil

        shutil.rmtree(outputs_path)
    dir_path.mkdir(parents=True, exist_ok=True)

    process = multiprocessing.Process(target=inference, args=(model_path, dummy_inputs, outputs_path, use_gpu))
    process.start()
    process.join()


def post_processing(outputs_path, outputs_path_other):
    # Compare outputs with e.g. fp16 and fp32
    record = {}
    if_close = {}

    import glob

    for filename in glob.glob(os.path.join(outputs_path, "*.tensorproto")):
        filename_other = os.path.join(outputs_path_other, Path(filename).name)
        if not os.path.exists(filename_other):
            continue
        with open(filename, "rb") as f:
            tensor = TensorProto()
            tensor.ParseFromString(f.read())
            array = numpy_helper.to_array(tensor)
            with open(filename_other, "rb") as f:  # noqa: PLW2901
                tensor_other = TensorProto()
                tensor_other.ParseFromString(f.read())
                array_other = numpy_helper.to_array(tensor_other)
                if array_other.size == 0:
                    continue
                diff = numpy.average(numpy.abs(array_other - array) / (numpy.abs(array_other) + 1e-6))
                if math.isnan(diff):
                    continue
                record[Path(filename).name.split(".")[0]] = diff
                if_close[Path(filename).name.split(".")[0]] = numpy.allclose(array, array_other, rtol=1e-04, atol=1e-04)

    results = ["Node\tDiff\tClose"]
    for k, v in sorted(record.items(), key=lambda x: x[1], reverse=True):
        results.append(f"{k}\t{v}\t{if_close[k]}")
    for line in results:
        print(line)


if __name__ == "__main__":
    # Below example shows how to use this helper to investigate parity issue of gpt-2 fp32 and fp16 onnx model
    # Please build ORT with --cmake_extra_defines onnxruntime_DEBUG_NODE_INPUTS_OUTPUTS=ON !!
    multiprocessing.set_start_method("spawn")

    # Generate Inputs
    sequence_length = 8
    past_sequence_length = 8
    batch_size = 5
    dummy_inputs_fp16 = Gpt2Helper.get_dummy_inputs(
        batch_size,
        past_sequence_length,
        sequence_length,
        12,
        768,
        12,
        50257,
        device=torch.device("cpu"),
        float16=True,
    )
    dummy_inputs_fp32 = dummy_inputs_fp16.to_fp32()

    # Get GPT-2 model from huggingface using convert_to_onnx.py
    os.system("python convert_to_onnx.py -m gpt2 --output gpt2_fp32.onnx -o -p fp32 --use_gpu")
    os.system("python convert_to_onnx.py -m gpt2 --output gpt2_fp16.onnx -o -p fp16 --use_gpu")

    # Specify the directory to dump the node's I/O
    outputs_path_fp32_gpu = "./fp32_gpu"
    outputs_path_fp16_gpu = "./fp16_gpu"
    generate_outputs_files("./gpt2_fp32.onnx", dummy_inputs_fp32, outputs_path_fp32_gpu, use_gpu=True)
    generate_outputs_files("./gpt2_fp16.onnx", dummy_inputs_fp16, outputs_path_fp16_gpu, use_gpu=True)

    # Compare each node's I/O value and sort based on average rtol
    post_processing(outputs_path_fp16_gpu, outputs_path_fp32_gpu)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\ops\common.py ===
"""
Boilerplate functions used in defining binary operations.
"""
from __future__ import annotations

from functools import wraps
from typing import (
    TYPE_CHECKING,
    Callable,
)

from pandas._libs.lib import item_from_zerodim
from pandas._libs.missing import is_matching_na

from pandas.core.dtypes.generic import (
    ABCIndex,
    ABCSeries,
)

if TYPE_CHECKING:
    from pandas._typing import F


def unpack_zerodim_and_defer(name: str) -> Callable[[F], F]:
    """
    Boilerplate for pandas conventions in arithmetic and comparison methods.

    Parameters
    ----------
    name : str

    Returns
    -------
    decorator
    """

    def wrapper(method: F) -> F:
        return _unpack_zerodim_and_defer(method, name)

    return wrapper


def _unpack_zerodim_and_defer(method, name: str):
    """
    Boilerplate for pandas conventions in arithmetic and comparison methods.

    Ensure method returns NotImplemented when operating against "senior"
    classes.  Ensure zero-dimensional ndarrays are always unpacked.

    Parameters
    ----------
    method : binary method
    name : str

    Returns
    -------
    method
    """
    stripped_name = name.removeprefix("__").removesuffix("__")
    is_cmp = stripped_name in {"eq", "ne", "lt", "le", "gt", "ge"}

    @wraps(method)
    def new_method(self, other):
        if is_cmp and isinstance(self, ABCIndex) and isinstance(other, ABCSeries):
            # For comparison ops, Index does *not* defer to Series
            pass
        else:
            prio = getattr(other, "__pandas_priority__", None)
            if prio is not None:
                if prio > self.__pandas_priority__:
                    # e.g. other is DataFrame while self is Index/Series/EA
                    return NotImplemented

        other = item_from_zerodim(other)

        return method(self, other)

    return new_method


def get_op_result_name(left, right):
    """
    Find the appropriate name to pin to an operation result.  This result
    should always be either an Index or a Series.

    Parameters
    ----------
    left : {Series, Index}
    right : object

    Returns
    -------
    name : object
        Usually a string
    """
    if isinstance(right, (ABCSeries, ABCIndex)):
        name = _maybe_match_name(left, right)
    else:
        name = left.name
    return name


def _maybe_match_name(a, b):
    """
    Try to find a name to attach to the result of an operation between
    a and b.  If only one of these has a `name` attribute, return that
    name.  Otherwise return a consensus name if they match or None if
    they have different names.

    Parameters
    ----------
    a : object
    b : object

    Returns
    -------
    name : str or None

    See Also
    --------
    pandas.core.common.consensus_name_attr
    """
    a_has = hasattr(a, "name")
    b_has = hasattr(b, "name")
    if a_has and b_has:
        try:
            if a.name == b.name:
                return a.name
            elif is_matching_na(a.name, b.name):
                # e.g. both are np.nan
                return a.name
            else:
                return None
        except TypeError:
            # pd.NA
            if is_matching_na(a.name, b.name):
                return a.name
            return None
        except ValueError:
            # e.g. np.int64(1) vs (np.int64(1), np.int64(2))
            return None
    elif a_has:
        return a.name
    elif b_has:
        return b.name
    return None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\commands\download.py ===
import logging
import os
from optparse import Values
from typing import List

from pip._internal.cli import cmdoptions
from pip._internal.cli.cmdoptions import make_target_python
from pip._internal.cli.req_command import RequirementCommand, with_cleanup
from pip._internal.cli.status_codes import SUCCESS
from pip._internal.operations.build.build_tracker import get_build_tracker
from pip._internal.req.req_install import check_legacy_setup_py_options
from pip._internal.utils.misc import ensure_dir, normalize_path, write_output
from pip._internal.utils.temp_dir import TempDirectory

logger = logging.getLogger(__name__)


class DownloadCommand(RequirementCommand):
    """
    Download packages from:

    - PyPI (and other indexes) using requirement specifiers.
    - VCS project urls.
    - Local project directories.
    - Local or remote source archives.

    pip also supports downloading from "requirements files", which provide
    an easy way to specify a whole environment to be downloaded.
    """

    usage = """
      %prog [options] <requirement specifier> [package-index-options] ...
      %prog [options] -r <requirements file> [package-index-options] ...
      %prog [options] <vcs project url> ...
      %prog [options] <local project path> ...
      %prog [options] <archive url/path> ..."""

    def add_options(self) -> None:
        self.cmd_opts.add_option(cmdoptions.constraints())
        self.cmd_opts.add_option(cmdoptions.requirements())
        self.cmd_opts.add_option(cmdoptions.no_deps())
        self.cmd_opts.add_option(cmdoptions.global_options())
        self.cmd_opts.add_option(cmdoptions.no_binary())
        self.cmd_opts.add_option(cmdoptions.only_binary())
        self.cmd_opts.add_option(cmdoptions.prefer_binary())
        self.cmd_opts.add_option(cmdoptions.src())
        self.cmd_opts.add_option(cmdoptions.pre())
        self.cmd_opts.add_option(cmdoptions.require_hashes())
        self.cmd_opts.add_option(cmdoptions.progress_bar())
        self.cmd_opts.add_option(cmdoptions.no_build_isolation())
        self.cmd_opts.add_option(cmdoptions.use_pep517())
        self.cmd_opts.add_option(cmdoptions.no_use_pep517())
        self.cmd_opts.add_option(cmdoptions.check_build_deps())
        self.cmd_opts.add_option(cmdoptions.ignore_requires_python())

        self.cmd_opts.add_option(
            "-d",
            "--dest",
            "--destination-dir",
            "--destination-directory",
            dest="download_dir",
            metavar="dir",
            default=os.curdir,
            help="Download packages into <dir>.",
        )

        cmdoptions.add_target_python_options(self.cmd_opts)

        index_opts = cmdoptions.make_option_group(
            cmdoptions.index_group,
            self.parser,
        )

        self.parser.insert_option_group(0, index_opts)
        self.parser.insert_option_group(0, self.cmd_opts)

    @with_cleanup
    def run(self, options: Values, args: List[str]) -> int:
        options.ignore_installed = True
        # editable doesn't really make sense for `pip download`, but the bowels
        # of the RequirementSet code require that property.
        options.editables = []

        cmdoptions.check_dist_restriction(options)

        options.download_dir = normalize_path(options.download_dir)
        ensure_dir(options.download_dir)

        session = self.get_default_session(options)

        target_python = make_target_python(options)
        finder = self._build_package_finder(
            options=options,
            session=session,
            target_python=target_python,
            ignore_requires_python=options.ignore_requires_python,
        )

        build_tracker = self.enter_context(get_build_tracker())

        directory = TempDirectory(
            delete=not options.no_clean,
            kind="download",
            globally_managed=True,
        )

        reqs = self.get_requirements(args, options, finder, session)
        check_legacy_setup_py_options(options, reqs)

        preparer = self.make_requirement_preparer(
            temp_build_dir=directory,
            options=options,
            build_tracker=build_tracker,
            session=session,
            finder=finder,
            download_dir=options.download_dir,
            use_user_site=False,
            verbosity=self.verbosity,
        )

        resolver = self.make_resolver(
            preparer=preparer,
            finder=finder,
            options=options,
            ignore_requires_python=options.ignore_requires_python,
            use_pep517=options.use_pep517,
            py_version_info=options.python_version,
        )

        self.trace_basic_info(finder)

        requirement_set = resolver.resolve(reqs, check_supported_wheels=True)

        downloaded: List[str] = []
        for req in requirement_set.requirements.values():
            if req.satisfied_by is None:
                assert req.name is not None
                preparer.save_linked_requirement(req)
                downloaded.append(req.name)

        preparer.prepare_linked_requirements_more(requirement_set.requirements.values())

        if downloaded:
            write_output("Successfully downloaded %s", " ".join(downloaded))

        return SUCCESS

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\cachecontrol\serialize.py ===
# SPDX-FileCopyrightText: 2015 Eric Larson
#
# SPDX-License-Identifier: Apache-2.0
from __future__ import annotations

import io
from typing import IO, TYPE_CHECKING, Any, Mapping, cast

from pip._vendor import msgpack
from pip._vendor.requests.structures import CaseInsensitiveDict
from pip._vendor.urllib3 import HTTPResponse

if TYPE_CHECKING:
    from pip._vendor.requests import PreparedRequest


class Serializer:
    serde_version = "4"

    def dumps(
        self,
        request: PreparedRequest,
        response: HTTPResponse,
        body: bytes | None = None,
    ) -> bytes:
        response_headers: CaseInsensitiveDict[str] = CaseInsensitiveDict(
            response.headers
        )

        if body is None:
            # When a body isn't passed in, we'll read the response. We
            # also update the response with a new file handler to be
            # sure it acts as though it was never read.
            body = response.read(decode_content=False)
            response._fp = io.BytesIO(body)  # type: ignore[assignment]
            response.length_remaining = len(body)

        data = {
            "response": {
                "body": body,  # Empty bytestring if body is stored separately
                "headers": {str(k): str(v) for k, v in response.headers.items()},
                "status": response.status,
                "version": response.version,
                "reason": str(response.reason),
                "decode_content": response.decode_content,
            }
        }

        # Construct our vary headers
        data["vary"] = {}
        if "vary" in response_headers:
            varied_headers = response_headers["vary"].split(",")
            for header in varied_headers:
                header = str(header).strip()
                header_value = request.headers.get(header, None)
                if header_value is not None:
                    header_value = str(header_value)
                data["vary"][header] = header_value

        return b",".join([f"cc={self.serde_version}".encode(), self.serialize(data)])

    def serialize(self, data: dict[str, Any]) -> bytes:
        return cast(bytes, msgpack.dumps(data, use_bin_type=True))

    def loads(
        self,
        request: PreparedRequest,
        data: bytes,
        body_file: IO[bytes] | None = None,
    ) -> HTTPResponse | None:
        # Short circuit if we've been given an empty set of data
        if not data:
            return None

        # Previous versions of this library supported other serialization
        # formats, but these have all been removed.
        if not data.startswith(f"cc={self.serde_version},".encode()):
            return None

        data = data[5:]
        return self._loads_v4(request, data, body_file)

    def prepare_response(
        self,
        request: PreparedRequest,
        cached: Mapping[str, Any],
        body_file: IO[bytes] | None = None,
    ) -> HTTPResponse | None:
        """Verify our vary headers match and construct a real urllib3
        HTTPResponse object.
        """
        # Special case the '*' Vary value as it means we cannot actually
        # determine if the cached response is suitable for this request.
        # This case is also handled in the controller code when creating
        # a cache entry, but is left here for backwards compatibility.
        if "*" in cached.get("vary", {}):
            return None

        # Ensure that the Vary headers for the cached response match our
        # request
        for header, value in cached.get("vary", {}).items():
            if request.headers.get(header, None) != value:
                return None

        body_raw = cached["response"].pop("body")

        headers: CaseInsensitiveDict[str] = CaseInsensitiveDict(
            data=cached["response"]["headers"]
        )
        if headers.get("transfer-encoding", "") == "chunked":
            headers.pop("transfer-encoding")

        cached["response"]["headers"] = headers

        try:
            body: IO[bytes]
            if body_file is None:
                body = io.BytesIO(body_raw)
            else:
                body = body_file
        except TypeError:
            # This can happen if cachecontrol serialized to v1 format (pickle)
            # using Python 2. A Python 2 str(byte string) will be unpickled as
            # a Python 3 str (unicode string), which will cause the above to
            # fail with:
            #
            #     TypeError: 'str' does not support the buffer interface
            body = io.BytesIO(body_raw.encode("utf8"))

        # Discard any `strict` parameter serialized by older version of cachecontrol.
        cached["response"].pop("strict", None)

        return HTTPResponse(body=body, preload_content=False, **cached["response"])

    def _loads_v4(
        self,
        request: PreparedRequest,
        data: bytes,
        body_file: IO[bytes] | None = None,
    ) -> HTTPResponse | None:
        try:
            cached = msgpack.loads(data, raw=False)
        except ValueError:
            return None

        return self.prepare_response(request, cached, body_file)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\h11\_writers.py ===
# Code to read HTTP data
#
# Strategy: each writer takes an event + a write-some-bytes function, which is
# calls.
#
# WRITERS is a dict describing how to pick a reader. It maps states to either:
# - a writer
# - or, for body writers, a dict of framin-dependent writer factories

from typing import Any, Callable, Dict, List, Tuple, Type, Union

from ._events import Data, EndOfMessage, Event, InformationalResponse, Request, Response
from ._headers import Headers
from ._state import CLIENT, IDLE, SEND_BODY, SEND_RESPONSE, SERVER
from ._util import LocalProtocolError, Sentinel

__all__ = ["WRITERS"]

Writer = Callable[[bytes], Any]


def write_headers(headers: Headers, write: Writer) -> None:
    # "Since the Host field-value is critical information for handling a
    # request, a user agent SHOULD generate Host as the first header field
    # following the request-line." - RFC 7230
    raw_items = headers._full_items
    for raw_name, name, value in raw_items:
        if name == b"host":
            write(b"%s: %s\r\n" % (raw_name, value))
    for raw_name, name, value in raw_items:
        if name != b"host":
            write(b"%s: %s\r\n" % (raw_name, value))
    write(b"\r\n")


def write_request(request: Request, write: Writer) -> None:
    if request.http_version != b"1.1":
        raise LocalProtocolError("I only send HTTP/1.1")
    write(b"%s %s HTTP/1.1\r\n" % (request.method, request.target))
    write_headers(request.headers, write)


# Shared between InformationalResponse and Response
def write_any_response(
    response: Union[InformationalResponse, Response], write: Writer
) -> None:
    if response.http_version != b"1.1":
        raise LocalProtocolError("I only send HTTP/1.1")
    status_bytes = str(response.status_code).encode("ascii")
    # We don't bother sending ascii status messages like "OK"; they're
    # optional and ignored by the protocol. (But the space after the numeric
    # status code is mandatory.)
    #
    # XX FIXME: could at least make an effort to pull out the status message
    # from stdlib's http.HTTPStatus table. Or maybe just steal their enums
    # (either by import or copy/paste). We already accept them as status codes
    # since they're of type IntEnum < int.
    write(b"HTTP/1.1 %s %s\r\n" % (status_bytes, response.reason))
    write_headers(response.headers, write)


class BodyWriter:
    def __call__(self, event: Event, write: Writer) -> None:
        if type(event) is Data:
            self.send_data(event.data, write)
        elif type(event) is EndOfMessage:
            self.send_eom(event.headers, write)
        else:  # pragma: no cover
            assert False

    def send_data(self, data: bytes, write: Writer) -> None:
        pass

    def send_eom(self, headers: Headers, write: Writer) -> None:
        pass


#
# These are all careful not to do anything to 'data' except call len(data) and
# write(data). This allows us to transparently pass-through funny objects,
# like placeholder objects referring to files on disk that will be sent via
# sendfile(2).
#
class ContentLengthWriter(BodyWriter):
    def __init__(self, length: int) -> None:
        self._length = length

    def send_data(self, data: bytes, write: Writer) -> None:
        self._length -= len(data)
        if self._length < 0:
            raise LocalProtocolError("Too much data for declared Content-Length")
        write(data)

    def send_eom(self, headers: Headers, write: Writer) -> None:
        if self._length != 0:
            raise LocalProtocolError("Too little data for declared Content-Length")
        if headers:
            raise LocalProtocolError("Content-Length and trailers don't mix")


class ChunkedWriter(BodyWriter):
    def send_data(self, data: bytes, write: Writer) -> None:
        # if we encoded 0-length data in the naive way, it would look like an
        # end-of-message.
        if not data:
            return
        write(b"%x\r\n" % len(data))
        write(data)
        write(b"\r\n")

    def send_eom(self, headers: Headers, write: Writer) -> None:
        write(b"0\r\n")
        write_headers(headers, write)


class Http10Writer(BodyWriter):
    def send_data(self, data: bytes, write: Writer) -> None:
        write(data)

    def send_eom(self, headers: Headers, write: Writer) -> None:
        if headers:
            raise LocalProtocolError("can't send trailers to HTTP/1.0 client")
        # no need to close the socket ourselves, that will be taken care of by
        # Connection: close machinery


WritersType = Dict[
    Union[Tuple[Type[Sentinel], Type[Sentinel]], Type[Sentinel]],
    Union[
        Dict[str, Type[BodyWriter]],
        Callable[[Union[InformationalResponse, Response], Writer], None],
        Callable[[Request, Writer], None],
    ],
]

WRITERS: WritersType = {
    (CLIENT, IDLE): write_request,
    (SERVER, IDLE): write_any_response,
    (SERVER, SEND_RESPONSE): write_any_response,
    SEND_BODY: {
        "chunked": ChunkedWriter,
        "content-length": ContentLengthWriter,
        "http/1.0": Http10Writer,
    },
}

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\packaging\licenses\__init__.py ===
#######################################################################################
#
# Adapted from:
#  https://github.com/pypa/hatch/blob/5352e44/backend/src/hatchling/licenses/parse.py
#
# MIT License
#
# Copyright (c) 2017-present Ofek Lev <oss@ofek.dev>
#
# Permission is hereby granted, free of charge, to any person obtaining a copy of this
# software and associated documentation files (the "Software"), to deal in the Software
# without restriction, including without limitation the rights to use, copy, modify,
# merge, publish, distribute, sublicense, and/or sell copies of the Software, and to
# permit persons to whom the Software is furnished to do so, subject to the following
# conditions:
#
# The above copyright notice and this permission notice shall be included in all copies
# or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A
# PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT
# HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF
# CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE
# OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
#
#
# With additional allowance of arbitrary `LicenseRef-` identifiers, not just
# `LicenseRef-Public-Domain` and `LicenseRef-Proprietary`.
#
#######################################################################################
from __future__ import annotations

import re
from typing import NewType, cast

from packaging.licenses._spdx import EXCEPTIONS, LICENSES

__all__ = [
    "InvalidLicenseExpression",
    "NormalizedLicenseExpression",
    "canonicalize_license_expression",
]

license_ref_allowed = re.compile("^[A-Za-z0-9.-]*$")

NormalizedLicenseExpression = NewType("NormalizedLicenseExpression", str)


class InvalidLicenseExpression(ValueError):
    """Raised when a license-expression string is invalid

    >>> canonicalize_license_expression("invalid")
    Traceback (most recent call last):
        ...
    packaging.licenses.InvalidLicenseExpression: Invalid license expression: 'invalid'
    """


def canonicalize_license_expression(
    raw_license_expression: str,
) -> NormalizedLicenseExpression:
    if not raw_license_expression:
        message = f"Invalid license expression: {raw_license_expression!r}"
        raise InvalidLicenseExpression(message)

    # Pad any parentheses so tokenization can be achieved by merely splitting on
    # whitespace.
    license_expression = raw_license_expression.replace("(", " ( ").replace(")", " ) ")
    licenseref_prefix = "LicenseRef-"
    license_refs = {
        ref.lower(): "LicenseRef-" + ref[len(licenseref_prefix) :]
        for ref in license_expression.split()
        if ref.lower().startswith(licenseref_prefix.lower())
    }

    # Normalize to lower case so we can look up licenses/exceptions
    # and so boolean operators are Python-compatible.
    license_expression = license_expression.lower()

    tokens = license_expression.split()

    # Rather than implementing boolean logic, we create an expression that Python can
    # parse. Everything that is not involved with the grammar itself is treated as
    # `False` and the expression should evaluate as such.
    python_tokens = []
    for token in tokens:
        if token not in {"or", "and", "with", "(", ")"}:
            python_tokens.append("False")
        elif token == "with":
            python_tokens.append("or")
        elif token == "(" and python_tokens and python_tokens[-1] not in {"or", "and"}:
            message = f"Invalid license expression: {raw_license_expression!r}"
            raise InvalidLicenseExpression(message)
        else:
            python_tokens.append(token)

    python_expression = " ".join(python_tokens)
    try:
        invalid = eval(python_expression, globals(), locals())
    except Exception:
        invalid = True

    if invalid is not False:
        message = f"Invalid license expression: {raw_license_expression!r}"
        raise InvalidLicenseExpression(message) from None

    # Take a final pass to check for unknown licenses/exceptions.
    normalized_tokens = []
    for token in tokens:
        if token in {"or", "and", "with", "(", ")"}:
            normalized_tokens.append(token.upper())
            continue

        if normalized_tokens and normalized_tokens[-1] == "WITH":
            if token not in EXCEPTIONS:
                message = f"Unknown license exception: {token!r}"
                raise InvalidLicenseExpression(message)

            normalized_tokens.append(EXCEPTIONS[token]["id"])
        else:
            if token.endswith("+"):
                final_token = token[:-1]
                suffix = "+"
            else:
                final_token = token
                suffix = ""

            if final_token.startswith("licenseref-"):
                if not license_ref_allowed.match(final_token):
                    message = f"Invalid licenseref: {final_token!r}"
                    raise InvalidLicenseExpression(message)
                normalized_tokens.append(license_refs[final_token] + suffix)
            else:
                if final_token not in LICENSES:
                    message = f"Unknown license: {final_token!r}"
                    raise InvalidLicenseExpression(message)
                normalized_tokens.append(LICENSES[final_token]["id"] + suffix)

    normalized_expression = " ".join(normalized_tokens)

    return cast(
        NormalizedLicenseExpression,
        normalized_expression.replace("( ", "(").replace(" )", ")"),
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\cachecontrol\caches\file_cache.py ===
# SPDX-FileCopyrightText: 2015 Eric Larson
#
# SPDX-License-Identifier: Apache-2.0
from __future__ import annotations

import hashlib
import os
import tempfile
from textwrap import dedent
from typing import IO, TYPE_CHECKING
from pathlib import Path

from pip._vendor.cachecontrol.cache import BaseCache, SeparateBodyBaseCache
from pip._vendor.cachecontrol.controller import CacheController

if TYPE_CHECKING:
    from datetime import datetime

    from filelock import BaseFileLock


class _FileCacheMixin:
    """Shared implementation for both FileCache variants."""

    def __init__(
        self,
        directory: str | Path,
        forever: bool = False,
        filemode: int = 0o0600,
        dirmode: int = 0o0700,
        lock_class: type[BaseFileLock] | None = None,
    ) -> None:
        try:
            if lock_class is None:
                from filelock import FileLock

                lock_class = FileLock
        except ImportError:
            notice = dedent(
                """
            NOTE: In order to use the FileCache you must have
            filelock installed. You can install it via pip:
              pip install cachecontrol[filecache]
            """
            )
            raise ImportError(notice)

        self.directory = directory
        self.forever = forever
        self.filemode = filemode
        self.dirmode = dirmode
        self.lock_class = lock_class

    @staticmethod
    def encode(x: str) -> str:
        return hashlib.sha224(x.encode()).hexdigest()

    def _fn(self, name: str) -> str:
        # NOTE: This method should not change as some may depend on it.
        #       See: https://github.com/ionrock/cachecontrol/issues/63
        hashed = self.encode(name)
        parts = list(hashed[:5]) + [hashed]
        return os.path.join(self.directory, *parts)

    def get(self, key: str) -> bytes | None:
        name = self._fn(key)
        try:
            with open(name, "rb") as fh:
                return fh.read()

        except FileNotFoundError:
            return None

    def set(
        self, key: str, value: bytes, expires: int | datetime | None = None
    ) -> None:
        name = self._fn(key)
        self._write(name, value)

    def _write(self, path: str, data: bytes) -> None:
        """
        Safely write the data to the given path.
        """
        # Make sure the directory exists
        dirname = os.path.dirname(path)
        os.makedirs(dirname, self.dirmode, exist_ok=True)

        with self.lock_class(path + ".lock"):
            # Write our actual file
            (fd, name) = tempfile.mkstemp(dir=dirname)
            try:
                os.write(fd, data)
            finally:
                os.close(fd)
            os.chmod(name, self.filemode)
            os.replace(name, path)

    def _delete(self, key: str, suffix: str) -> None:
        name = self._fn(key) + suffix
        if not self.forever:
            try:
                os.remove(name)
            except FileNotFoundError:
                pass


class FileCache(_FileCacheMixin, BaseCache):
    """
    Traditional FileCache: body is stored in memory, so not suitable for large
    downloads.
    """

    def delete(self, key: str) -> None:
        self._delete(key, "")


class SeparateBodyFileCache(_FileCacheMixin, SeparateBodyBaseCache):
    """
    Memory-efficient FileCache: body is stored in a separate file, reducing
    peak memory usage.
    """

    def get_body(self, key: str) -> IO[bytes] | None:
        name = self._fn(key) + ".body"
        try:
            return open(name, "rb")
        except FileNotFoundError:
            return None

    def set_body(self, key: str, body: bytes) -> None:
        name = self._fn(key) + ".body"
        self._write(name, body)

    def delete(self, key: str) -> None:
        self._delete(key, "")
        self._delete(key, ".body")


def url_to_file_path(url: str, filecache: FileCache) -> str:
    """Return the file cache path based on the URL.

    This does not ensure the file exists!
    """
    key = CacheController.cache_url(url)
    return filecache._fn(key)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\packaging\licenses\__init__.py ===
#######################################################################################
#
# Adapted from:
#  https://github.com/pypa/hatch/blob/5352e44/backend/src/hatchling/licenses/parse.py
#
# MIT License
#
# Copyright (c) 2017-present Ofek Lev <oss@ofek.dev>
#
# Permission is hereby granted, free of charge, to any person obtaining a copy of this
# software and associated documentation files (the "Software"), to deal in the Software
# without restriction, including without limitation the rights to use, copy, modify,
# merge, publish, distribute, sublicense, and/or sell copies of the Software, and to
# permit persons to whom the Software is furnished to do so, subject to the following
# conditions:
#
# The above copyright notice and this permission notice shall be included in all copies
# or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A
# PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT
# HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF
# CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE
# OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
#
#
# With additional allowance of arbitrary `LicenseRef-` identifiers, not just
# `LicenseRef-Public-Domain` and `LicenseRef-Proprietary`.
#
#######################################################################################
from __future__ import annotations

import re
from typing import NewType, cast

from pip._vendor.packaging.licenses._spdx import EXCEPTIONS, LICENSES

__all__ = [
    "InvalidLicenseExpression",
    "NormalizedLicenseExpression",
    "canonicalize_license_expression",
]

license_ref_allowed = re.compile("^[A-Za-z0-9.-]*$")

NormalizedLicenseExpression = NewType("NormalizedLicenseExpression", str)


class InvalidLicenseExpression(ValueError):
    """Raised when a license-expression string is invalid

    >>> canonicalize_license_expression("invalid")
    Traceback (most recent call last):
        ...
    packaging.licenses.InvalidLicenseExpression: Invalid license expression: 'invalid'
    """


def canonicalize_license_expression(
    raw_license_expression: str,
) -> NormalizedLicenseExpression:
    if not raw_license_expression:
        message = f"Invalid license expression: {raw_license_expression!r}"
        raise InvalidLicenseExpression(message)

    # Pad any parentheses so tokenization can be achieved by merely splitting on
    # whitespace.
    license_expression = raw_license_expression.replace("(", " ( ").replace(")", " ) ")
    licenseref_prefix = "LicenseRef-"
    license_refs = {
        ref.lower(): "LicenseRef-" + ref[len(licenseref_prefix) :]
        for ref in license_expression.split()
        if ref.lower().startswith(licenseref_prefix.lower())
    }

    # Normalize to lower case so we can look up licenses/exceptions
    # and so boolean operators are Python-compatible.
    license_expression = license_expression.lower()

    tokens = license_expression.split()

    # Rather than implementing boolean logic, we create an expression that Python can
    # parse. Everything that is not involved with the grammar itself is treated as
    # `False` and the expression should evaluate as such.
    python_tokens = []
    for token in tokens:
        if token not in {"or", "and", "with", "(", ")"}:
            python_tokens.append("False")
        elif token == "with":
            python_tokens.append("or")
        elif token == "(" and python_tokens and python_tokens[-1] not in {"or", "and"}:
            message = f"Invalid license expression: {raw_license_expression!r}"
            raise InvalidLicenseExpression(message)
        else:
            python_tokens.append(token)

    python_expression = " ".join(python_tokens)
    try:
        invalid = eval(python_expression, globals(), locals())
    except Exception:
        invalid = True

    if invalid is not False:
        message = f"Invalid license expression: {raw_license_expression!r}"
        raise InvalidLicenseExpression(message) from None

    # Take a final pass to check for unknown licenses/exceptions.
    normalized_tokens = []
    for token in tokens:
        if token in {"or", "and", "with", "(", ")"}:
            normalized_tokens.append(token.upper())
            continue

        if normalized_tokens and normalized_tokens[-1] == "WITH":
            if token not in EXCEPTIONS:
                message = f"Unknown license exception: {token!r}"
                raise InvalidLicenseExpression(message)

            normalized_tokens.append(EXCEPTIONS[token]["id"])
        else:
            if token.endswith("+"):
                final_token = token[:-1]
                suffix = "+"
            else:
                final_token = token
                suffix = ""

            if final_token.startswith("licenseref-"):
                if not license_ref_allowed.match(final_token):
                    message = f"Invalid licenseref: {final_token!r}"
                    raise InvalidLicenseExpression(message)
                normalized_tokens.append(license_refs[final_token] + suffix)
            else:
                if final_token not in LICENSES:
                    message = f"Unknown license: {final_token!r}"
                    raise InvalidLicenseExpression(message)
                normalized_tokens.append(LICENSES[final_token]["id"] + suffix)

    normalized_expression = " ".join(normalized_tokens)

    return cast(
        NormalizedLicenseExpression,
        normalized_expression.replace("( ", "(").replace(" )", ")"),
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\sql\__init__.py ===
# sql/__init__.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
from typing import Any
from typing import TYPE_CHECKING

from ._typing import ColumnExpressionArgument as ColumnExpressionArgument
from ._typing import NotNullable as NotNullable
from ._typing import Nullable as Nullable
from .base import Executable as Executable
from .compiler import COLLECT_CARTESIAN_PRODUCTS as COLLECT_CARTESIAN_PRODUCTS
from .compiler import FROM_LINTING as FROM_LINTING
from .compiler import NO_LINTING as NO_LINTING
from .compiler import WARN_LINTING as WARN_LINTING
from .ddl import BaseDDLElement as BaseDDLElement
from .ddl import DDL as DDL
from .ddl import DDLElement as DDLElement
from .ddl import ExecutableDDLElement as ExecutableDDLElement
from .expression import Alias as Alias
from .expression import alias as alias
from .expression import all_ as all_
from .expression import and_ as and_
from .expression import any_ as any_
from .expression import asc as asc
from .expression import between as between
from .expression import bindparam as bindparam
from .expression import case as case
from .expression import cast as cast
from .expression import ClauseElement as ClauseElement
from .expression import collate as collate
from .expression import column as column
from .expression import ColumnCollection as ColumnCollection
from .expression import ColumnElement as ColumnElement
from .expression import CompoundSelect as CompoundSelect
from .expression import cte as cte
from .expression import Delete as Delete
from .expression import delete as delete
from .expression import desc as desc
from .expression import distinct as distinct
from .expression import except_ as except_
from .expression import except_all as except_all
from .expression import exists as exists
from .expression import extract as extract
from .expression import false as false
from .expression import False_ as False_
from .expression import FromClause as FromClause
from .expression import func as func
from .expression import funcfilter as funcfilter
from .expression import Insert as Insert
from .expression import insert as insert
from .expression import intersect as intersect
from .expression import intersect_all as intersect_all
from .expression import Join as Join
from .expression import join as join
from .expression import label as label
from .expression import LABEL_STYLE_DEFAULT as LABEL_STYLE_DEFAULT
from .expression import (
    LABEL_STYLE_DISAMBIGUATE_ONLY as LABEL_STYLE_DISAMBIGUATE_ONLY,
)
from .expression import LABEL_STYLE_NONE as LABEL_STYLE_NONE
from .expression import (
    LABEL_STYLE_TABLENAME_PLUS_COL as LABEL_STYLE_TABLENAME_PLUS_COL,
)
from .expression import lambda_stmt as lambda_stmt
from .expression import LambdaElement as LambdaElement
from .expression import lateral as lateral
from .expression import literal as literal
from .expression import literal_column as literal_column
from .expression import modifier as modifier
from .expression import not_ as not_
from .expression import null as null
from .expression import nulls_first as nulls_first
from .expression import nulls_last as nulls_last
from .expression import nullsfirst as nullsfirst
from .expression import nullslast as nullslast
from .expression import or_ as or_
from .expression import outerjoin as outerjoin
from .expression import outparam as outparam
from .expression import over as over
from .expression import quoted_name as quoted_name
from .expression import Select as Select
from .expression import select as select
from .expression import Selectable as Selectable
from .expression import SelectLabelStyle as SelectLabelStyle
from .expression import SQLColumnExpression as SQLColumnExpression
from .expression import StatementLambdaElement as StatementLambdaElement
from .expression import Subquery as Subquery
from .expression import table as table
from .expression import TableClause as TableClause
from .expression import TableSample as TableSample
from .expression import tablesample as tablesample
from .expression import text as text
from .expression import true as true
from .expression import True_ as True_
from .expression import try_cast as try_cast
from .expression import tuple_ as tuple_
from .expression import type_coerce as type_coerce
from .expression import union as union
from .expression import union_all as union_all
from .expression import Update as Update
from .expression import update as update
from .expression import Values as Values
from .expression import values as values
from .expression import within_group as within_group
from .visitors import ClauseVisitor as ClauseVisitor


def __go(lcls: Any) -> None:
    from .. import util as _sa_util

    from . import base
    from . import coercions
    from . import elements
    from . import lambdas
    from . import selectable
    from . import schema
    from . import traversals
    from . import type_api

    if not TYPE_CHECKING:
        base.coercions = elements.coercions = coercions
        base.elements = elements
        base.type_api = type_api
        coercions.elements = elements
        coercions.lambdas = lambdas
        coercions.schema = schema
        coercions.selectable = selectable

    from .annotation import _prepare_annotations
    from .annotation import Annotated
    from .elements import AnnotatedColumnElement
    from .elements import ClauseList
    from .selectable import AnnotatedFromClause

    _prepare_annotations(ColumnElement, AnnotatedColumnElement)
    _prepare_annotations(FromClause, AnnotatedFromClause)
    _prepare_annotations(ClauseList, Annotated)

    _sa_util.preloaded.import_prefix("sqlalchemy.sql")


__go(locals())

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\parsing\latex\lark\latex_parser.py ===
import os
import logging
import re
from pathlib import Path

from sympy.external import import_module
from sympy.parsing.latex.lark.transformer import TransformToSymPyExpr

_lark = import_module("lark")


class LarkLaTeXParser:
    r"""Class for converting input `\mathrm{\LaTeX}` strings into SymPy Expressions.
    It holds all the necessary internal data for doing so, and exposes hooks for
    customizing its behavior.

    Parameters
    ==========

    print_debug_output : bool, optional

        If set to ``True``, prints debug output to the logger. Defaults to ``False``.

    transform : bool, optional

        If set to ``True``, the class runs the Transformer class on the parse tree
        generated by running ``Lark.parse`` on the input string. Defaults to ``True``.

        Setting it to ``False`` can help with debugging the `\mathrm{\LaTeX}` grammar.

    grammar_file : str, optional

        The path to the grammar file that the parser should use. If set to ``None``,
        it uses the default grammar, which is in ``grammar/latex.lark``, relative to
        the ``sympy/parsing/latex/lark/`` directory.

    transformer : str, optional

        The name of the Transformer class to use. If set to ``None``, it uses the
        default transformer class, which is :py:func:`TransformToSymPyExpr`.

    """
    def __init__(self, print_debug_output=False, transform=True, grammar_file=None, transformer=None):
        grammar_dir_path = os.path.join(os.path.dirname(__file__), "grammar/")

        if grammar_file is None:
            latex_grammar = Path(os.path.join(grammar_dir_path, "latex.lark")).read_text(encoding="utf-8")
        else:
            latex_grammar = Path(grammar_file).read_text(encoding="utf-8")

        self.parser = _lark.Lark(
            latex_grammar,
            source_path=grammar_dir_path,
            parser="earley",
            start="latex_string",
            lexer="auto",
            ambiguity="explicit",
            propagate_positions=False,
            maybe_placeholders=False,
            keep_all_tokens=True)

        self.print_debug_output = print_debug_output
        self.transform_expr = transform

        if transformer is None:
            self.transformer = TransformToSymPyExpr()
        else:
            self.transformer = transformer()

    def doparse(self, s: str):
        if self.print_debug_output:
            _lark.logger.setLevel(logging.DEBUG)

        parse_tree = self.parser.parse(s)

        if not self.transform_expr:
            # exit early and return the parse tree
            _lark.logger.debug("expression = %s", s)
            _lark.logger.debug(parse_tree)
            _lark.logger.debug(parse_tree.pretty())
            return parse_tree

        if self.print_debug_output:
            # print this stuff before attempting to run the transformer
            _lark.logger.debug("expression = %s", s)
            # print the `parse_tree` variable
            _lark.logger.debug(parse_tree.pretty())

        sympy_expression = self.transformer.transform(parse_tree)

        if self.print_debug_output:
            _lark.logger.debug("SymPy expression = %s", sympy_expression)

        return sympy_expression


if _lark is not None:
    _lark_latex_parser = LarkLaTeXParser()


def parse_latex_lark(s: str):
    """
    Experimental LaTeX parser using Lark.

    This function is still under development and its API may change with the
    next releases of SymPy.
    """
    if _lark is None:
        raise ImportError("Lark is probably not installed")
    return _lark_latex_parser.doparse(s)


def _pretty_print_lark_trees(tree, indent=0, show_expr=True):
    if isinstance(tree, _lark.Token):
        return tree.value

    data = str(tree.data)

    is_expr = data.startswith("expression")

    if is_expr:
        data = re.sub(r"^expression", "E", data)

    is_ambig = (data == "_ambig")

    if is_ambig:
        new_indent = indent + 2
    else:
        new_indent = indent

    output = ""
    show_node = not is_expr or show_expr

    if show_node:
        output += str(data) + "("

    if is_ambig:
        output += "\n" + "\n".join([" " * new_indent + _pretty_print_lark_trees(i, new_indent, show_expr) for i in tree.children])
    else:
        output += ",".join([_pretty_print_lark_trees(i, new_indent, show_expr) for i in tree.children])

    if show_node:
        output += ")"

    return output

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_constant_fold.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

from logging import getLogger

from fusion_base import Fusion
from fusion_utils import NumpyHelper
from onnx_model import OnnxModel

logger = getLogger(__name__)


class FusionConstantFold(Fusion):
    def __init__(self, model: OnnxModel):
        super().__init__(model, "", ["Transpose"])
        self.count = 0

    def apply(self):
        super().apply()
        if self.count > 0:
            logger.info(f"Constant Folded: {self.count}")

    def fuse(self, node, input_name_to_nodes, output_name_to_node):
        """
        Apply multiple fusions on Transpose nodes that can be constant folded.
        """
        self.fuse_1(node, input_name_to_nodes, output_name_to_node)
        self.fuse_2(node, input_name_to_nodes, output_name_to_node)

    def fuse_1(self, node, input_name_to_nodes, output_name_to_node):
        """
        Constant fold any initializer data representing a MatMul's
        weights that are stored in a Transpose op

        Ex: Transpose --> Gemm or Transpose --> MatMul
        """
        # Check if Transpose node only has one input and one output
        if len(node.input) != 1 or len(node.output) != 1:
            logger.debug("fuse_constant_fold: node has more than one input or output")
            return

        # Check if input is initializer data
        proto = self.model.get_initializer(node.input[0])
        if proto is None:
            logger.debug("fuse_constant_fold: failed to identify initializer input")
            return

        # Check that all nodes using input are Transpose ops that also only use the initializer data as input
        skip = False
        for child_node in input_name_to_nodes[node.input[0]]:
            if not (child_node.op_type == "Transpose" and len(node.input) == 1):
                skip = True
                break
        if skip:
            logger.debug("fuse_constant_fold: other non-Transpose nodes use the initializer")
            return

        # Check that all nodes using output are Gemm or MatMul ops
        for child_node in input_name_to_nodes[node.output[0]]:
            if not (child_node.op_type == "Gemm" or child_node.op_type == "MatMul"):
                skip = True
                break
        if skip:
            logger.debug("fuse_constant_fold: other non-Gemm and non-MatMul nodes use the transposed data")
            return

        # Check if initializer data is 2D
        weight = NumpyHelper.to_array(proto)
        if len(weight.shape) != 2:
            logger.debug("fuse_constant_fold: shape of initializer data is not 2D")
            return

        # Remove old TensorProto and add new TensorProto while re-using same name
        name = proto.name
        dtype = proto.data_type
        self.remove_initializer(proto)
        self.add_initializer(
            name=name,
            data_type=dtype,
            dims=[weight.shape[1], weight.shape[0]],
            vals=weight.T,
        )

        # Update weights input to be the initializer name and not
        # the output of the Transpose op
        for child_node in input_name_to_nodes[node.output[0]]:
            for i in range(len(child_node.input)):
                if child_node.input[i] == node.output[0]:
                    child_node.input[i] = node.input[0]

                    if child_node.op_type == "Gemm" and (i == 0 or i == 1):
                        # Ensure that transA/transB is set to 0 in Gemm
                        key = "transA" if i == 0 else "transB"
                        for j, attr_key in enumerate(child_node.attribute):
                            if attr_key.name == key:
                                child_node.attribute[j].i = 0

        # Add node to list of nodes to remove
        self.nodes_to_remove.append(node)
        self.count += 1

    def fuse_2(self, node, input_name_to_nodes, output_name_to_node):
        """
        Constant fold any Transpose --> Transpose ops since the root input
        is the final result

        Ex: root_input --> Transpose --> Transpose --> next_node to root_input --> next_node
        """
        # Check if Transpose node only has one input and one output
        if len(node.input) != 1 or len(node.output) != 1:
            logger.debug("fuse_constant_fold: node has more than one input or output")
            return

        # Check if parent node is Transpose node with only one input and one output
        parent_node = self.model.match_parent(node, "Transpose", 0)
        if parent_node is None:
            logger.debug("fuse_constant_fold: failed to identify parent Transpose node")
            return
        if len(parent_node.input) != 1 or len(parent_node.output) != 1:
            logger.debug("fuse_constant_fold: parent node has more than one input or output")
            return

        node_perm = node.attribute[0].ints
        parent_node_perm = parent_node.attribute[0].ints

        if node_perm != parent_node_perm:
            logger.debug("fuse_constant_fold: Transpose node permutations aren't identical")
            return

        # For nodes that use output of child Transpose node as an input,
        # replace that input with root_input
        root_input = parent_node.input[0]
        output_nodes = input_name_to_nodes[node.output[0]]
        for output_node in output_nodes:
            for i, input_ in enumerate(output_node.input):
                if input_ == node.output[0]:
                    output_node.input[i] = root_input

        # Add node to list of nodes to remove
        self.nodes_to_remove.append(node)
        self.nodes_to_remove.append(parent_node)
        self.count += 1

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\chart\bar_chart.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Typed,
    Bool,
    Integer,
    Sequence,
    Alias,
)
from openpyxl.descriptors.excel import ExtensionList
from openpyxl.descriptors.nested import (
    NestedNoneSet,
    NestedSet,
    NestedBool,
    NestedInteger,
    NestedMinMax,
)

from .descriptors import (
    NestedGapAmount,
    NestedOverlap,
)
from ._chart import ChartBase
from ._3d import _3DBase
from .axis import TextAxis, NumericAxis, SeriesAxis, ChartLines
from .shapes import GraphicalProperties
from .series import Series
from .legend import Legend
from .label import DataLabelList


class _BarChartBase(ChartBase):

    barDir = NestedSet(values=(['bar', 'col']))
    type = Alias("barDir")
    grouping = NestedSet(values=(['percentStacked', 'clustered', 'standard',
                                  'stacked']))
    varyColors = NestedBool(nested=True, allow_none=True)
    ser = Sequence(expected_type=Series, allow_none=True)
    dLbls = Typed(expected_type=DataLabelList, allow_none=True)
    dataLabels = Alias("dLbls")

    __elements__ = ('barDir', 'grouping', 'varyColors', 'ser', 'dLbls')

    _series_type = "bar"

    def __init__(self,
                 barDir="col",
                 grouping="clustered",
                 varyColors=None,
                 ser=(),
                 dLbls=None,
                 **kw
                ):
        self.barDir = barDir
        self.grouping = grouping
        self.varyColors = varyColors
        self.ser = ser
        self.dLbls = dLbls
        super().__init__(**kw)


class BarChart(_BarChartBase):

    tagname = "barChart"

    barDir = _BarChartBase.barDir
    grouping = _BarChartBase.grouping
    varyColors = _BarChartBase.varyColors
    ser = _BarChartBase.ser
    dLbls = _BarChartBase.dLbls

    gapWidth = NestedGapAmount()
    overlap = NestedOverlap()
    serLines = Typed(expected_type=ChartLines, allow_none=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    # chart properties actually used by containing classes
    x_axis = Typed(expected_type=TextAxis)
    y_axis = Typed(expected_type=NumericAxis)

    __elements__ = _BarChartBase.__elements__ + ('gapWidth', 'overlap', 'serLines', 'axId')

    def __init__(self,
                 gapWidth=150,
                 overlap=None,
                 serLines=None,
                 extLst=None,
                 **kw
                ):
        self.gapWidth = gapWidth
        self.overlap = overlap
        self.serLines = serLines
        self.x_axis = TextAxis()
        self.y_axis = NumericAxis()
        self.legend = Legend()
        super().__init__(**kw)


class BarChart3D(_BarChartBase, _3DBase):

    tagname = "bar3DChart"

    barDir = _BarChartBase.barDir
    grouping = _BarChartBase.grouping
    varyColors = _BarChartBase.varyColors
    ser = _BarChartBase.ser
    dLbls = _BarChartBase.dLbls

    view3D = _3DBase.view3D
    floor = _3DBase.floor
    sideWall = _3DBase.sideWall
    backWall = _3DBase.backWall

    gapWidth = NestedGapAmount()
    gapDepth = NestedGapAmount()
    shape = NestedNoneSet(values=(['cone', 'coneToMax', 'box', 'cylinder', 'pyramid', 'pyramidToMax']))
    serLines = Typed(expected_type=ChartLines, allow_none=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    x_axis = Typed(expected_type=TextAxis)
    y_axis = Typed(expected_type=NumericAxis)
    z_axis = Typed(expected_type=SeriesAxis, allow_none=True)

    __elements__ = _BarChartBase.__elements__ + ('gapWidth', 'gapDepth', 'shape', 'serLines', 'axId')

    def __init__(self,
                 gapWidth=150,
                 gapDepth=150,
                 shape=None,
                 serLines=None,
                 extLst=None,
                 **kw
                ):
        self.gapWidth = gapWidth
        self.gapDepth = gapDepth
        self.shape = shape
        self.serLines = serLines
        self.x_axis = TextAxis()
        self.y_axis = NumericAxis()
        self.z_axis = SeriesAxis()

        super(BarChart3D, self).__init__(**kw)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\drawing\connector.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Typed,
    Bool,
    Integer,
    String,
    Alias,
)
from openpyxl.descriptors.excel import ExtensionList as OfficeArtExtensionList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText

from .properties import (
    NonVisualDrawingProps,
    NonVisualDrawingShapeProps,
)
from .geometry import ShapeStyle

class Connection(Serialisable):

    id = Integer()
    idx = Integer()

    def __init__(self,
                 id=None,
                 idx=None,
                ):
        self.id = id
        self.idx = idx


class ConnectorLocking(Serialisable):

    extLst = Typed(expected_type=OfficeArtExtensionList, allow_none=True)

    def __init__(self,
                 extLst=None,
                ):
        self.extLst = extLst


class NonVisualConnectorProperties(Serialisable):

    cxnSpLocks = Typed(expected_type=ConnectorLocking, allow_none=True)
    stCxn = Typed(expected_type=Connection, allow_none=True)
    endCxn = Typed(expected_type=Connection, allow_none=True)
    extLst = Typed(expected_type=OfficeArtExtensionList, allow_none=True)

    def __init__(self,
                 cxnSpLocks=None,
                 stCxn=None,
                 endCxn=None,
                 extLst=None,
                ):
        self.cxnSpLocks = cxnSpLocks
        self.stCxn = stCxn
        self.endCxn = endCxn
        self.extLst = extLst


class ConnectorNonVisual(Serialisable):

    cNvPr = Typed(expected_type=NonVisualDrawingProps, )
    cNvCxnSpPr = Typed(expected_type=NonVisualConnectorProperties, )

    __elements__ = ("cNvPr", "cNvCxnSpPr",)

    def __init__(self,
                 cNvPr=None,
                 cNvCxnSpPr=None,
                ):
        self.cNvPr = cNvPr
        self.cNvCxnSpPr = cNvCxnSpPr


class ConnectorShape(Serialisable):

    tagname = "cxnSp"

    nvCxnSpPr = Typed(expected_type=ConnectorNonVisual)
    spPr = Typed(expected_type=GraphicalProperties)
    style = Typed(expected_type=ShapeStyle, allow_none=True)
    macro = String(allow_none=True)
    fPublished = Bool(allow_none=True)

    def __init__(self,
                 nvCxnSpPr=None,
                 spPr=None,
                 style=None,
                 macro=None,
                 fPublished=None,
                 ):
        self.nvCxnSpPr = nvCxnSpPr
        self.spPr = spPr
        self.style = style
        self.macro = macro
        self.fPublished = fPublished


class ShapeMeta(Serialisable):

    tagname = "nvSpPr"

    cNvPr = Typed(expected_type=NonVisualDrawingProps)
    cNvSpPr = Typed(expected_type=NonVisualDrawingShapeProps)

    def __init__(self, cNvPr=None, cNvSpPr=None):
        self.cNvPr = cNvPr
        self.cNvSpPr = cNvSpPr


class Shape(Serialisable):

    macro = String(allow_none=True)
    textlink = String(allow_none=True)
    fPublished = Bool(allow_none=True)
    fLocksText = Bool(allow_none=True)
    nvSpPr = Typed(expected_type=ShapeMeta, allow_none=True)
    meta = Alias("nvSpPr")
    spPr = Typed(expected_type=GraphicalProperties)
    graphicalProperties = Alias("spPr")
    style = Typed(expected_type=ShapeStyle, allow_none=True)
    txBody = Typed(expected_type=RichText, allow_none=True)

    def __init__(self,
                 macro=None,
                 textlink=None,
                 fPublished=None,
                 fLocksText=None,
                 nvSpPr=None,
                 spPr=None,
                 style=None,
                 txBody=None,
                ):
        self.macro = macro
        self.textlink = textlink
        self.fPublished = fPublished
        self.fLocksText = fLocksText
        self.nvSpPr = nvSpPr
        self.spPr = spPr
        self.style = style
        self.txBody = txBody

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\drawing\line.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Typed,
    Integer,
    MinMax,
    NoneSet,
    Alias,
    Sequence
)

from openpyxl.descriptors.nested import (
    NestedInteger,
    NestedNoneSet,
    EmptyTag,
)
from openpyxl.xml.constants import DRAWING_NS

from .colors import ColorChoiceDescriptor
from .fill import GradientFillProperties, PatternFillProperties
from openpyxl.descriptors.excel import ExtensionList as OfficeArtExtensionList

"""
Line elements from drawing main schema
"""


class LineEndProperties(Serialisable):

    tagname = "end"
    namespace = DRAWING_NS

    type = NoneSet(values=(['none', 'triangle', 'stealth', 'diamond', 'oval', 'arrow']))
    w = NoneSet(values=(['sm', 'med', 'lg']))
    len = NoneSet(values=(['sm', 'med', 'lg']))

    def __init__(self,
                 type=None,
                 w=None,
                 len=None,
                ):
        self.type = type
        self.w = w
        self.len = len


class DashStop(Serialisable):

    tagname = "ds"
    namespace = DRAWING_NS

    d = Integer()
    length = Alias('d')
    sp = Integer()
    space = Alias('sp')

    def __init__(self,
                 d=0,
                 sp=0,
                ):
        self.d = d
        self.sp = sp


class DashStopList(Serialisable):

    ds = Sequence(expected_type=DashStop, allow_none=True)

    def __init__(self,
                 ds=None,
                ):
        self.ds = ds


class LineProperties(Serialisable):

    tagname = "ln"
    namespace = DRAWING_NS

    w = MinMax(min=0, max=20116800, allow_none=True) # EMU
    width = Alias('w')
    cap = NoneSet(values=(['rnd', 'sq', 'flat']))
    cmpd = NoneSet(values=(['sng', 'dbl', 'thickThin', 'thinThick', 'tri']))
    algn = NoneSet(values=(['ctr', 'in']))

    noFill = EmptyTag()
    solidFill = ColorChoiceDescriptor()
    gradFill = Typed(expected_type=GradientFillProperties, allow_none=True)
    pattFill = Typed(expected_type=PatternFillProperties, allow_none=True)

    prstDash = NestedNoneSet(values=(['solid', 'dot', 'dash', 'lgDash', 'dashDot',
                       'lgDashDot', 'lgDashDotDot', 'sysDash', 'sysDot', 'sysDashDot',
                       'sysDashDotDot']), namespace=namespace)
    dashStyle = Alias('prstDash')

    custDash = Typed(expected_type=DashStop, allow_none=True)

    round = EmptyTag()
    bevel = EmptyTag()
    miter = NestedInteger(allow_none=True, attribute="lim")

    headEnd = Typed(expected_type=LineEndProperties, allow_none=True)
    tailEnd = Typed(expected_type=LineEndProperties, allow_none=True)
    extLst = Typed(expected_type=OfficeArtExtensionList, allow_none=True)

    __elements__ = ('noFill', 'solidFill', 'gradFill', 'pattFill',
                    'prstDash', 'custDash', 'round', 'bevel', 'miter', 'headEnd', 'tailEnd')

    def __init__(self,
                 w=None,
                 cap=None,
                 cmpd=None,
                 algn=None,
                 noFill=None,
                 solidFill=None,
                 gradFill=None,
                 pattFill=None,
                 prstDash=None,
                 custDash=None,
                 round=None,
                 bevel=None,
                 miter=None,
                 headEnd=None,
                 tailEnd=None,
                 extLst=None,
                ):
        self.w = w
        self.cap = cap
        self.cmpd = cmpd
        self.algn = algn
        self.noFill = noFill
        self.solidFill = solidFill
        self.gradFill = gradFill
        self.pattFill = pattFill
        if prstDash is None:
            prstDash = "solid"
        self.prstDash = prstDash
        self.custDash = custDash
        self.round = round
        self.bevel = bevel
        self.miter = miter
        self.headEnd = headEnd
        self.tailEnd = tailEnd

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\drawing\picture.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.xml.constants import DRAWING_NS

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Typed,
    Bool,
    String,
    Alias,
)
from openpyxl.descriptors.excel import ExtensionList as OfficeArtExtensionList

from openpyxl.chart.shapes import GraphicalProperties

from .fill import BlipFillProperties
from .properties import NonVisualDrawingProps
from .geometry import ShapeStyle


class PictureLocking(Serialisable):

    tagname = "picLocks"
    namespace = DRAWING_NS

    # Using attribute group AG_Locking
    noCrop = Bool(allow_none=True)
    noGrp = Bool(allow_none=True)
    noSelect = Bool(allow_none=True)
    noRot = Bool(allow_none=True)
    noChangeAspect = Bool(allow_none=True)
    noMove = Bool(allow_none=True)
    noResize = Bool(allow_none=True)
    noEditPoints = Bool(allow_none=True)
    noAdjustHandles = Bool(allow_none=True)
    noChangeArrowheads = Bool(allow_none=True)
    noChangeShapeType = Bool(allow_none=True)
    extLst = Typed(expected_type=OfficeArtExtensionList, allow_none=True)

    __elements__ = ()

    def __init__(self,
                 noCrop=None,
                 noGrp=None,
                 noSelect=None,
                 noRot=None,
                 noChangeAspect=None,
                 noMove=None,
                 noResize=None,
                 noEditPoints=None,
                 noAdjustHandles=None,
                 noChangeArrowheads=None,
                 noChangeShapeType=None,
                 extLst=None,
                ):
        self.noCrop = noCrop
        self.noGrp = noGrp
        self.noSelect = noSelect
        self.noRot = noRot
        self.noChangeAspect = noChangeAspect
        self.noMove = noMove
        self.noResize = noResize
        self.noEditPoints = noEditPoints
        self.noAdjustHandles = noAdjustHandles
        self.noChangeArrowheads = noChangeArrowheads
        self.noChangeShapeType = noChangeShapeType


class NonVisualPictureProperties(Serialisable):

    tagname = "cNvPicPr"

    preferRelativeResize = Bool(allow_none=True)
    picLocks = Typed(expected_type=PictureLocking, allow_none=True)
    extLst = Typed(expected_type=OfficeArtExtensionList, allow_none=True)

    __elements__ = ("picLocks",)

    def __init__(self,
                 preferRelativeResize=None,
                 picLocks=None,
                 extLst=None,
                ):
        self.preferRelativeResize = preferRelativeResize
        self.picLocks = picLocks


class PictureNonVisual(Serialisable):

    tagname = "nvPicPr"

    cNvPr = Typed(expected_type=NonVisualDrawingProps, )
    cNvPicPr = Typed(expected_type=NonVisualPictureProperties, )

    __elements__ = ("cNvPr", "cNvPicPr")

    def __init__(self,
                 cNvPr=None,
                 cNvPicPr=None,
                ):
        if cNvPr is None:
            cNvPr = NonVisualDrawingProps(id=0, name="Image 1", descr="Name of file")
        self.cNvPr = cNvPr
        if cNvPicPr is None:
            cNvPicPr = NonVisualPictureProperties()
        self.cNvPicPr = cNvPicPr




class PictureFrame(Serialisable):

    tagname = "pic"

    macro = String(allow_none=True)
    fPublished = Bool(allow_none=True)
    nvPicPr = Typed(expected_type=PictureNonVisual, )
    blipFill = Typed(expected_type=BlipFillProperties, )
    spPr = Typed(expected_type=GraphicalProperties, )
    graphicalProperties = Alias('spPr')
    style = Typed(expected_type=ShapeStyle, allow_none=True)

    __elements__ = ("nvPicPr", "blipFill", "spPr", "style")

    def __init__(self,
                 macro=None,
                 fPublished=None,
                 nvPicPr=None,
                 blipFill=None,
                 spPr=None,
                 style=None,
                ):
        self.macro = macro
        self.fPublished = fPublished
        if nvPicPr is None:
            nvPicPr = PictureNonVisual()
        self.nvPicPr = nvPicPr
        if blipFill is None:
            blipFill = BlipFillProperties()
        self.blipFill = blipFill
        if spPr is None:
            spPr = GraphicalProperties()
        self.spPr = spPr
        self.style = style

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\cli\progress_bars.py ===
import functools
import sys
from typing import Callable, Generator, Iterable, Iterator, Optional, Tuple, TypeVar

from pip._vendor.rich.progress import (
    BarColumn,
    DownloadColumn,
    FileSizeColumn,
    MofNCompleteColumn,
    Progress,
    ProgressColumn,
    SpinnerColumn,
    TextColumn,
    TimeElapsedColumn,
    TimeRemainingColumn,
    TransferSpeedColumn,
)

from pip._internal.cli.spinners import RateLimiter
from pip._internal.req.req_install import InstallRequirement
from pip._internal.utils.logging import get_console, get_indentation

T = TypeVar("T")
ProgressRenderer = Callable[[Iterable[T]], Iterator[T]]


def _rich_download_progress_bar(
    iterable: Iterable[bytes],
    *,
    bar_type: str,
    size: Optional[int],
    initial_progress: Optional[int] = None,
) -> Generator[bytes, None, None]:
    assert bar_type == "on", "This should only be used in the default mode."

    if not size:
        total = float("inf")
        columns: Tuple[ProgressColumn, ...] = (
            TextColumn("[progress.description]{task.description}"),
            SpinnerColumn("line", speed=1.5),
            FileSizeColumn(),
            TransferSpeedColumn(),
            TimeElapsedColumn(),
        )
    else:
        total = size
        columns = (
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            DownloadColumn(),
            TransferSpeedColumn(),
            TextColumn("eta"),
            TimeRemainingColumn(),
        )

    progress = Progress(*columns, refresh_per_second=5)
    task_id = progress.add_task(" " * (get_indentation() + 2), total=total)
    if initial_progress is not None:
        progress.update(task_id, advance=initial_progress)
    with progress:
        for chunk in iterable:
            yield chunk
            progress.update(task_id, advance=len(chunk))


def _rich_install_progress_bar(
    iterable: Iterable[InstallRequirement], *, total: int
) -> Iterator[InstallRequirement]:
    columns = (
        TextColumn("{task.fields[indent]}"),
        BarColumn(),
        MofNCompleteColumn(),
        TextColumn("{task.description}"),
    )
    console = get_console()

    bar = Progress(*columns, refresh_per_second=6, console=console, transient=True)
    # Hiding the progress bar at initialization forces a refresh cycle to occur
    # until the bar appears, avoiding very short flashes.
    task = bar.add_task("", total=total, indent=" " * get_indentation(), visible=False)
    with bar:
        for req in iterable:
            bar.update(task, description=rf"\[{req.name}]", visible=True)
            yield req
            bar.advance(task)


def _raw_progress_bar(
    iterable: Iterable[bytes],
    *,
    size: Optional[int],
    initial_progress: Optional[int] = None,
) -> Generator[bytes, None, None]:
    def write_progress(current: int, total: int) -> None:
        sys.stdout.write(f"Progress {current} of {total}\n")
        sys.stdout.flush()

    current = initial_progress or 0
    total = size or 0
    rate_limiter = RateLimiter(0.25)

    write_progress(current, total)
    for chunk in iterable:
        current += len(chunk)
        if rate_limiter.ready() or current == total:
            write_progress(current, total)
            rate_limiter.reset()
        yield chunk


def get_download_progress_renderer(
    *, bar_type: str, size: Optional[int] = None, initial_progress: Optional[int] = None
) -> ProgressRenderer[bytes]:
    """Get an object that can be used to render the download progress.

    Returns a callable, that takes an iterable to "wrap".
    """
    if bar_type == "on":
        return functools.partial(
            _rich_download_progress_bar,
            bar_type=bar_type,
            size=size,
            initial_progress=initial_progress,
        )
    elif bar_type == "raw":
        return functools.partial(
            _raw_progress_bar,
            size=size,
            initial_progress=initial_progress,
        )
    else:
        return iter  # no-op, when passed an iterator


def get_install_progress_renderer(
    *, bar_type: str, total: int
) -> ProgressRenderer[InstallRequirement]:
    """Get an object that can be used to render the install progress.
    Returns a callable, that takes an iterable to "wrap".
    """
    if bar_type == "on":
        return functools.partial(_rich_install_progress_bar, total=total)
    else:
        return iter

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\platformdirs\macos.py ===
"""macOS."""

from __future__ import annotations

import os.path
import sys
from typing import TYPE_CHECKING

from .api import PlatformDirsABC

if TYPE_CHECKING:
    from pathlib import Path


class MacOS(PlatformDirsABC):
    """
    Platform directories for the macOS operating system.

    Follows the guidance from
    `Apple documentation <https://developer.apple.com/library/archive/documentation/FileManagement/Conceptual/FileSystemProgrammingGuide/MacOSXDirectories/MacOSXDirectories.html>`_.
    Makes use of the `appname <platformdirs.api.PlatformDirsABC.appname>`,
    `version <platformdirs.api.PlatformDirsABC.version>`,
    `ensure_exists <platformdirs.api.PlatformDirsABC.ensure_exists>`.

    """

    @property
    def user_data_dir(self) -> str:
        """:return: data directory tied to the user, e.g. ``~/Library/Application Support/$appname/$version``"""
        return self._append_app_name_and_version(os.path.expanduser("~/Library/Application Support"))  # noqa: PTH111

    @property
    def site_data_dir(self) -> str:
        """
        :return: data directory shared by users, e.g. ``/Library/Application Support/$appname/$version``.
          If we're using a Python binary managed by `Homebrew <https://brew.sh>`_, the directory
          will be under the Homebrew prefix, e.g. ``/opt/homebrew/share/$appname/$version``.
          If `multipath <platformdirs.api.PlatformDirsABC.multipath>` is enabled, and we're in Homebrew,
          the response is a multi-path string separated by ":", e.g.
          ``/opt/homebrew/share/$appname/$version:/Library/Application Support/$appname/$version``
        """
        is_homebrew = sys.prefix.startswith("/opt/homebrew")
        path_list = [self._append_app_name_and_version("/opt/homebrew/share")] if is_homebrew else []
        path_list.append(self._append_app_name_and_version("/Library/Application Support"))
        if self.multipath:
            return os.pathsep.join(path_list)
        return path_list[0]

    @property
    def site_data_path(self) -> Path:
        """:return: data path shared by users. Only return the first item, even if ``multipath`` is set to ``True``"""
        return self._first_item_as_path_if_multipath(self.site_data_dir)

    @property
    def user_config_dir(self) -> str:
        """:return: config directory tied to the user, same as `user_data_dir`"""
        return self.user_data_dir

    @property
    def site_config_dir(self) -> str:
        """:return: config directory shared by the users, same as `site_data_dir`"""
        return self.site_data_dir

    @property
    def user_cache_dir(self) -> str:
        """:return: cache directory tied to the user, e.g. ``~/Library/Caches/$appname/$version``"""
        return self._append_app_name_and_version(os.path.expanduser("~/Library/Caches"))  # noqa: PTH111

    @property
    def site_cache_dir(self) -> str:
        """
        :return: cache directory shared by users, e.g. ``/Library/Caches/$appname/$version``.
          If we're using a Python binary managed by `Homebrew <https://brew.sh>`_, the directory
          will be under the Homebrew prefix, e.g. ``/opt/homebrew/var/cache/$appname/$version``.
          If `multipath <platformdirs.api.PlatformDirsABC.multipath>` is enabled, and we're in Homebrew,
          the response is a multi-path string separated by ":", e.g.
          ``/opt/homebrew/var/cache/$appname/$version:/Library/Caches/$appname/$version``
        """
        is_homebrew = sys.prefix.startswith("/opt/homebrew")
        path_list = [self._append_app_name_and_version("/opt/homebrew/var/cache")] if is_homebrew else []
        path_list.append(self._append_app_name_and_version("/Library/Caches"))
        if self.multipath:
            return os.pathsep.join(path_list)
        return path_list[0]

    @property
    def site_cache_path(self) -> Path:
        """:return: cache path shared by users. Only return the first item, even if ``multipath`` is set to ``True``"""
        return self._first_item_as_path_if_multipath(self.site_cache_dir)

    @property
    def user_state_dir(self) -> str:
        """:return: state directory tied to the user, same as `user_data_dir`"""
        return self.user_data_dir

    @property
    def user_log_dir(self) -> str:
        """:return: log directory tied to the user, e.g. ``~/Library/Logs/$appname/$version``"""
        return self._append_app_name_and_version(os.path.expanduser("~/Library/Logs"))  # noqa: PTH111

    @property
    def user_documents_dir(self) -> str:
        """:return: documents directory tied to the user, e.g. ``~/Documents``"""
        return os.path.expanduser("~/Documents")  # noqa: PTH111

    @property
    def user_downloads_dir(self) -> str:
        """:return: downloads directory tied to the user, e.g. ``~/Downloads``"""
        return os.path.expanduser("~/Downloads")  # noqa: PTH111

    @property
    def user_pictures_dir(self) -> str:
        """:return: pictures directory tied to the user, e.g. ``~/Pictures``"""
        return os.path.expanduser("~/Pictures")  # noqa: PTH111

    @property
    def user_videos_dir(self) -> str:
        """:return: videos directory tied to the user, e.g. ``~/Movies``"""
        return os.path.expanduser("~/Movies")  # noqa: PTH111

    @property
    def user_music_dir(self) -> str:
        """:return: music directory tied to the user, e.g. ``~/Music``"""
        return os.path.expanduser("~/Music")  # noqa: PTH111

    @property
    def user_desktop_dir(self) -> str:
        """:return: desktop directory tied to the user, e.g. ``~/Desktop``"""
        return os.path.expanduser("~/Desktop")  # noqa: PTH111

    @property
    def user_runtime_dir(self) -> str:
        """:return: runtime directory tied to the user, e.g. ``~/Library/Caches/TemporaryItems/$appname/$version``"""
        return self._append_app_name_and_version(os.path.expanduser("~/Library/Caches/TemporaryItems"))  # noqa: PTH111

    @property
    def site_runtime_dir(self) -> str:
        """:return: runtime directory shared by users, same as `user_runtime_dir`"""
        return self.user_runtime_dir


__all__ = [
    "MacOS",
]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\plotting\pygletplot\plot_window.py ===
from time import perf_counter


import pyglet.gl as pgl

from sympy.plotting.pygletplot.managed_window import ManagedWindow
from sympy.plotting.pygletplot.plot_camera import PlotCamera
from sympy.plotting.pygletplot.plot_controller import PlotController


class PlotWindow(ManagedWindow):

    def __init__(self, plot, antialiasing=True, ortho=False,
                 invert_mouse_zoom=False, linewidth=1.5, caption="SymPy Plot",
                 **kwargs):
        """
        Named Arguments
        ===============

        antialiasing = True
            True OR False
        ortho = False
            True OR False
        invert_mouse_zoom = False
            True OR False
        """
        self.plot = plot

        self.camera = None
        self._calculating = False

        self.antialiasing = antialiasing
        self.ortho = ortho
        self.invert_mouse_zoom = invert_mouse_zoom
        self.linewidth = linewidth
        self.title = caption
        self.last_caption_update = 0
        self.caption_update_interval = 0.2
        self.drawing_first_object = True

        super().__init__(**kwargs)

    def setup(self):
        self.camera = PlotCamera(self, ortho=self.ortho)
        self.controller = PlotController(self,
                invert_mouse_zoom=self.invert_mouse_zoom)
        self.push_handlers(self.controller)

        pgl.glClearColor(1.0, 1.0, 1.0, 0.0)
        pgl.glClearDepth(1.0)

        pgl.glDepthFunc(pgl.GL_LESS)
        pgl.glEnable(pgl.GL_DEPTH_TEST)

        pgl.glEnable(pgl.GL_LINE_SMOOTH)
        pgl.glShadeModel(pgl.GL_SMOOTH)
        pgl.glLineWidth(self.linewidth)

        pgl.glEnable(pgl.GL_BLEND)
        pgl.glBlendFunc(pgl.GL_SRC_ALPHA, pgl.GL_ONE_MINUS_SRC_ALPHA)

        if self.antialiasing:
            pgl.glHint(pgl.GL_LINE_SMOOTH_HINT, pgl.GL_NICEST)
            pgl.glHint(pgl.GL_POLYGON_SMOOTH_HINT, pgl.GL_NICEST)

        self.camera.setup_projection()

    def on_resize(self, w, h):
        super().on_resize(w, h)
        if self.camera is not None:
            self.camera.setup_projection()

    def update(self, dt):
        self.controller.update(dt)

    def draw(self):
        self.plot._render_lock.acquire()
        self.camera.apply_transformation()

        calc_verts_pos, calc_verts_len = 0, 0
        calc_cverts_pos, calc_cverts_len = 0, 0

        should_update_caption = (perf_counter() - self.last_caption_update >
                                 self.caption_update_interval)

        if len(self.plot._functions.values()) == 0:
            self.drawing_first_object = True

        iterfunctions = iter(self.plot._functions.values())

        for r in iterfunctions:
            if self.drawing_first_object:
                self.camera.set_rot_preset(r.default_rot_preset)
                self.drawing_first_object = False

            pgl.glPushMatrix()
            r._draw()
            pgl.glPopMatrix()

            # might as well do this while we are
            # iterating and have the lock rather
            # than locking and iterating twice
            # per frame:

            if should_update_caption:
                try:
                    if r.calculating_verts:
                        calc_verts_pos += r.calculating_verts_pos
                        calc_verts_len += r.calculating_verts_len
                    if r.calculating_cverts:
                        calc_cverts_pos += r.calculating_cverts_pos
                        calc_cverts_len += r.calculating_cverts_len
                except ValueError:
                    pass

        for r in self.plot._pobjects:
            pgl.glPushMatrix()
            r._draw()
            pgl.glPopMatrix()

        if should_update_caption:
            self.update_caption(calc_verts_pos, calc_verts_len,
                                calc_cverts_pos, calc_cverts_len)
            self.last_caption_update = perf_counter()

        if self.plot._screenshot:
            self.plot._screenshot._execute_saving()

        self.plot._render_lock.release()

    def update_caption(self, calc_verts_pos, calc_verts_len,
            calc_cverts_pos, calc_cverts_len):
        caption = self.title
        if calc_verts_len or calc_cverts_len:
            caption += " (calculating"
            if calc_verts_len > 0:
                p = (calc_verts_pos / calc_verts_len) * 100
                caption += " vertices %i%%" % (p)
            if calc_cverts_len > 0:
                p = (calc_cverts_pos / calc_cverts_len) * 100
                caption += " colors %i%%" % (p)
            caption += ")"
        if self.caption != caption:
            self.set_caption(caption)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\sets\handlers\issubset.py ===
from sympy.core.singleton import S
from sympy.core.symbol import Symbol
from sympy.core.logic import fuzzy_and, fuzzy_bool, fuzzy_not, fuzzy_or
from sympy.core.relational import Eq
from sympy.sets.sets import FiniteSet, Interval, Set, Union, ProductSet
from sympy.sets.fancysets import Complexes, Reals, Range, Rationals
from sympy.multipledispatch import Dispatcher


_inf_sets = [S.Naturals, S.Naturals0, S.Integers, S.Rationals, S.Reals, S.Complexes]


is_subset_sets = Dispatcher('is_subset_sets')


@is_subset_sets.register(Set, Set)
def _(a, b):
    return None

@is_subset_sets.register(Interval, Interval)
def _(a, b):
    # This is correct but can be made more comprehensive...
    if fuzzy_bool(a.start < b.start):
        return False
    if fuzzy_bool(a.end > b.end):
        return False
    if (b.left_open and not a.left_open and fuzzy_bool(Eq(a.start, b.start))):
        return False
    if (b.right_open and not a.right_open and fuzzy_bool(Eq(a.end, b.end))):
        return False

@is_subset_sets.register(Interval, FiniteSet)
def _(a_interval, b_fs):
    # An Interval can only be a subset of a finite set if it is finite
    # which can only happen if it has zero measure.
    if fuzzy_not(a_interval.measure.is_zero):
        return False

@is_subset_sets.register(Interval, Union)
def _(a_interval, b_u):
    if all(isinstance(s, (Interval, FiniteSet)) for s in b_u.args):
        intervals = [s for s in b_u.args if isinstance(s, Interval)]
        if all(fuzzy_bool(a_interval.start < s.start) for s in intervals):
            return False
        if all(fuzzy_bool(a_interval.end > s.end) for s in intervals):
            return False
        if a_interval.measure.is_nonzero:
            no_overlap = lambda s1, s2: fuzzy_or([
                    fuzzy_bool(s1.end <= s2.start),
                    fuzzy_bool(s1.start >= s2.end),
                    ])
            if all(no_overlap(s, a_interval) for s in intervals):
                return False

@is_subset_sets.register(Range, Range)
def _(a, b):
    if a.step == b.step == 1:
        return fuzzy_and([fuzzy_bool(a.start >= b.start),
                          fuzzy_bool(a.stop <= b.stop)])

@is_subset_sets.register(Range, Interval)
def _(a_range, b_interval):
    if a_range.step.is_positive:
        if b_interval.left_open and a_range.inf.is_finite:
            cond_left = a_range.inf > b_interval.left
        else:
            cond_left = a_range.inf >= b_interval.left
        if b_interval.right_open and a_range.sup.is_finite:
            cond_right = a_range.sup < b_interval.right
        else:
            cond_right = a_range.sup <= b_interval.right
        return fuzzy_and([cond_left, cond_right])

@is_subset_sets.register(Range, FiniteSet)
def _(a_range, b_finiteset):
    try:
        a_size = a_range.size
    except ValueError:
        # symbolic Range of unknown size
        return None
    if a_size > len(b_finiteset):
        return False
    elif any(arg.has(Symbol) for arg in a_range.args):
        return fuzzy_and(b_finiteset.contains(x) for x in a_range)
    else:
        # Checking A \ B == EmptySet is more efficient than repeated naive
        # membership checks on an arbitrary FiniteSet.
        a_set = set(a_range)
        b_remaining = len(b_finiteset)
        # Symbolic expressions and numbers of unknown type (integer or not) are
        # all counted as "candidates", i.e. *potentially* matching some a in
        # a_range.
        cnt_candidate = 0
        for b in b_finiteset:
            if b.is_Integer:
                a_set.discard(b)
            elif fuzzy_not(b.is_integer):
                pass
            else:
                cnt_candidate += 1
            b_remaining -= 1
            if len(a_set) > b_remaining + cnt_candidate:
                return False
            if len(a_set) == 0:
                return True
        return None

@is_subset_sets.register(Interval, Range)
def _(a_interval, b_range):
    if a_interval.measure.is_extended_nonzero:
        return False

@is_subset_sets.register(Interval, Rationals)
def _(a_interval, b_rationals):
    if a_interval.measure.is_extended_nonzero:
        return False

@is_subset_sets.register(Range, Complexes)
def _(a, b):
    return True

@is_subset_sets.register(Complexes, Interval)
def _(a, b):
    return False

@is_subset_sets.register(Complexes, Range)
def _(a, b):
    return False

@is_subset_sets.register(Complexes, Rationals)
def _(a, b):
    return False

@is_subset_sets.register(Rationals, Reals)
def _(a, b):
    return True

@is_subset_sets.register(Rationals, Range)
def _(a, b):
    return False

@is_subset_sets.register(ProductSet, FiniteSet)
def _(a_ps, b_fs):
    return fuzzy_and(b_fs.contains(x) for x in a_ps)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_windows_pipes.py ===
from __future__ import annotations

import sys
from typing import TYPE_CHECKING

from . import _core
from ._abc import ReceiveStream, SendStream
from ._core._windows_cffi import _handle, kernel32, raise_winerror
from ._util import ConflictDetector, final

assert sys.platform == "win32" or not TYPE_CHECKING

# XX TODO: don't just make this up based on nothing.
DEFAULT_RECEIVE_SIZE = 65536


# See the comments on _unix_pipes._FdHolder for discussion of why we set the
# handle to -1 when it's closed.
class _HandleHolder:
    def __init__(self, handle: int) -> None:
        self.handle = -1
        if not isinstance(handle, int):
            raise TypeError("handle must be an int")
        self.handle = handle
        _core.register_with_iocp(self.handle)

    @property
    def closed(self) -> bool:
        return self.handle == -1

    def close(self) -> None:
        if self.closed:
            return
        handle = self.handle
        self.handle = -1
        if not kernel32.CloseHandle(_handle(handle)):
            raise_winerror()

    def __del__(self) -> None:
        self.close()


@final
class PipeSendStream(SendStream):
    """Represents a send stream over a Windows named pipe that has been
    opened in OVERLAPPED mode.
    """

    def __init__(self, handle: int) -> None:
        self._handle_holder = _HandleHolder(handle)
        self._conflict_detector = ConflictDetector(
            "another task is currently using this pipe",
        )

    async def send_all(self, data: bytes) -> None:
        with self._conflict_detector:
            if self._handle_holder.closed:
                raise _core.ClosedResourceError("this pipe is already closed")

            if not data:
                await _core.checkpoint()
                return

            try:
                written = await _core.write_overlapped(self._handle_holder.handle, data)
            except BrokenPipeError as ex:
                raise _core.BrokenResourceError from ex
            # By my reading of MSDN, this assert is guaranteed to pass so long
            # as the pipe isn't in nonblocking mode, but... let's just
            # double-check.
            assert written == len(data)

    async def wait_send_all_might_not_block(self) -> None:
        with self._conflict_detector:
            if self._handle_holder.closed:
                raise _core.ClosedResourceError("This pipe is already closed")

            # not implemented yet, and probably not needed
            await _core.checkpoint()

    def close(self) -> None:
        self._handle_holder.close()

    async def aclose(self) -> None:
        self.close()
        await _core.checkpoint()


@final
class PipeReceiveStream(ReceiveStream):
    """Represents a receive stream over an os.pipe object."""

    def __init__(self, handle: int) -> None:
        self._handle_holder = _HandleHolder(handle)
        self._conflict_detector = ConflictDetector(
            "another task is currently using this pipe",
        )

    async def receive_some(self, max_bytes: int | None = None) -> bytes:
        with self._conflict_detector:
            if self._handle_holder.closed:
                raise _core.ClosedResourceError("this pipe is already closed")

            if max_bytes is None:
                max_bytes = DEFAULT_RECEIVE_SIZE
            else:
                if not isinstance(max_bytes, int):
                    raise TypeError("max_bytes must be integer >= 1")
                if max_bytes < 1:
                    raise ValueError("max_bytes must be integer >= 1")

            buffer = bytearray(max_bytes)
            try:
                size = await _core.readinto_overlapped(
                    self._handle_holder.handle,
                    buffer,
                )
            except BrokenPipeError:
                if self._handle_holder.closed:
                    raise _core.ClosedResourceError(
                        "another task closed this pipe",
                    ) from None

                # Windows raises BrokenPipeError on one end of a pipe
                # whenever the other end closes, regardless of direction.
                # Convert this to the Unix behavior of returning EOF to the
                # reader when the writer closes.
                #
                # And since we're not raising an exception, we have to
                # checkpoint. But readinto_overlapped did raise an exception,
                # so it might not have checkpointed for us. So we have to
                # checkpoint manually.
                await _core.checkpoint()
                return b""
            else:
                del buffer[size:]
                return buffer

    def close(self) -> None:
        self._handle_holder.close()

    async def aclose(self) -> None:
        self.close()
        await _core.checkpoint()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\f2py\common_rules.py ===
"""
Build common block mechanism for f2py2e.

Copyright 1999 -- 2011 Pearu Peterson all rights reserved.
Copyright 2011 -- present NumPy Developers.
Permission to use, modify, and distribute this software is given under the
terms of the NumPy License

NO WARRANTY IS EXPRESSED OR IMPLIED.  USE AT YOUR OWN RISK.
"""
from . import __version__

f2py_version = __version__.version

from . import capi_maps, func2subr
from .auxfuncs import getuseblocks, hasbody, hascommon, hasnote, isintent_hide, outmess
from .crackfortran import rmbadname


def findcommonblocks(block, top=1):
    ret = []
    if hascommon(block):
        for key, value in block['common'].items():
            vars_ = {v: block['vars'][v] for v in value}
            ret.append((key, value, vars_))
    elif hasbody(block):
        for b in block['body']:
            ret = ret + findcommonblocks(b, 0)
    if top:
        tret = []
        names = []
        for t in ret:
            if t[0] not in names:
                names.append(t[0])
                tret.append(t)
        return tret
    return ret


def buildhooks(m):
    ret = {'commonhooks': [], 'initcommonhooks': [],
           'docs': ['"COMMON blocks:\\n"']}
    fwrap = ['']

    def fadd(line, s=fwrap):
        s[0] = f'{s[0]}\n      {line}'
    chooks = ['']

    def cadd(line, s=chooks):
        s[0] = f'{s[0]}\n{line}'
    ihooks = ['']

    def iadd(line, s=ihooks):
        s[0] = f'{s[0]}\n{line}'
    doc = ['']

    def dadd(line, s=doc):
        s[0] = f'{s[0]}\n{line}'
    for (name, vnames, vars) in findcommonblocks(m):
        lower_name = name.lower()
        hnames, inames = [], []
        for n in vnames:
            if isintent_hide(vars[n]):
                hnames.append(n)
            else:
                inames.append(n)
        if hnames:
            outmess('\t\tConstructing COMMON block support for "%s"...\n\t\t  %s\n\t\t  Hidden: %s\n' % (
                name, ','.join(inames), ','.join(hnames)))
        else:
            outmess('\t\tConstructing COMMON block support for "%s"...\n\t\t  %s\n' % (
                name, ','.join(inames)))
        fadd(f'subroutine f2pyinit{name}(setupfunc)')
        for usename in getuseblocks(m):
            fadd(f'use {usename}')
        fadd('external setupfunc')
        for n in vnames:
            fadd(func2subr.var2fixfortran(vars, n))
        if name == '_BLNK_':
            fadd(f"common {','.join(vnames)}")
        else:
            fadd(f"common /{name}/ {','.join(vnames)}")
        fadd(f"call setupfunc({','.join(inames)})")
        fadd('end\n')
        cadd('static FortranDataDef f2py_%s_def[] = {' % (name))
        idims = []
        for n in inames:
            ct = capi_maps.getctype(vars[n])
            elsize = capi_maps.get_elsize(vars[n])
            at = capi_maps.c2capi_map[ct]
            dm = capi_maps.getarrdims(n, vars[n])
            if dm['dims']:
                idims.append(f"({dm['dims']})")
            else:
                idims.append('')
            dms = dm['dims'].strip()
            if not dms:
                dms = '-1'
            cadd('\t{\"%s\",%s,{{%s}},%s, %s},'
                 % (n, dm['rank'], dms, at, elsize))
        cadd('\t{NULL}\n};')
        inames1 = rmbadname(inames)
        inames1_tps = ','.join(['char *' + s for s in inames1])
        cadd('static void f2py_setup_%s(%s) {' % (name, inames1_tps))
        cadd('\tint i_f2py=0;')
        for n in inames1:
            cadd(f'\tf2py_{name}_def[i_f2py++].data = {n};')
        cadd('}')
        if '_' in lower_name:
            F_FUNC = 'F_FUNC_US'
        else:
            F_FUNC = 'F_FUNC'
        cadd('extern void %s(f2pyinit%s,F2PYINIT%s)(void(*)(%s));'
             % (F_FUNC, lower_name, name.upper(),
                ','.join(['char*'] * len(inames1))))
        cadd('static void f2py_init_%s(void) {' % name)
        cadd('\t%s(f2pyinit%s,F2PYINIT%s)(f2py_setup_%s);'
             % (F_FUNC, lower_name, name.upper(), name))
        cadd('}\n')
        iadd(f'\ttmp = PyFortranObject_New(f2py_{name}_def,f2py_init_{name});')
        iadd('\tif (tmp == NULL) return NULL;')
        iadd(f'\tif (F2PyDict_SetItemString(d, "{name}", tmp) == -1) return NULL;')
        iadd('\tPy_DECREF(tmp);')
        tname = name.replace('_', '\\_')
        dadd('\\subsection{Common block \\texttt{%s}}\n' % (tname))
        dadd('\\begin{description}')
        for n in inames:
            dadd('\\item[]{{}\\verb@%s@{}}' %
                 (capi_maps.getarrdocsign(n, vars[n])))
            if hasnote(vars[n]):
                note = vars[n]['note']
                if isinstance(note, list):
                    note = '\n'.join(note)
                dadd(f'--- {note}')
        dadd('\\end{description}')
        ret['docs'].append(
            f"\"\t/{name}/ {','.join(map(lambda v, d: v + d, inames, idims))}\\n\"")
    ret['commonhooks'] = chooks
    ret['initcommonhooks'] = ihooks
    ret['latexdoc'] = doc[0]
    if len(ret['docs']) <= 1:
        ret['docs'] = ''
    return ret, fwrap[0]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\computation\engines.py ===
"""
Engine classes for :func:`~pandas.eval`
"""
from __future__ import annotations

import abc
from typing import TYPE_CHECKING

from pandas.errors import NumExprClobberingError

from pandas.core.computation.align import (
    align_terms,
    reconstruct_object,
)
from pandas.core.computation.ops import (
    MATHOPS,
    REDUCTIONS,
)

from pandas.io.formats import printing

if TYPE_CHECKING:
    from pandas.core.computation.expr import Expr

_ne_builtins = frozenset(MATHOPS + REDUCTIONS)


def _check_ne_builtin_clash(expr: Expr) -> None:
    """
    Attempt to prevent foot-shooting in a helpful way.

    Parameters
    ----------
    expr : Expr
        Terms can contain
    """
    names = expr.names
    overlap = names & _ne_builtins

    if overlap:
        s = ", ".join([repr(x) for x in overlap])
        raise NumExprClobberingError(
            f'Variables in expression "{expr}" overlap with builtins: ({s})'
        )


class AbstractEngine(metaclass=abc.ABCMeta):
    """Object serving as a base class for all engines."""

    has_neg_frac = False

    def __init__(self, expr) -> None:
        self.expr = expr
        self.aligned_axes = None
        self.result_type = None

    def convert(self) -> str:
        """
        Convert an expression for evaluation.

        Defaults to return the expression as a string.
        """
        return printing.pprint_thing(self.expr)

    def evaluate(self) -> object:
        """
        Run the engine on the expression.

        This method performs alignment which is necessary no matter what engine
        is being used, thus its implementation is in the base class.

        Returns
        -------
        object
            The result of the passed expression.
        """
        if not self._is_aligned:
            self.result_type, self.aligned_axes = align_terms(self.expr.terms)

        # make sure no names in resolvers and locals/globals clash
        res = self._evaluate()
        return reconstruct_object(
            self.result_type, res, self.aligned_axes, self.expr.terms.return_type
        )

    @property
    def _is_aligned(self) -> bool:
        return self.aligned_axes is not None and self.result_type is not None

    @abc.abstractmethod
    def _evaluate(self):
        """
        Return an evaluated expression.

        Parameters
        ----------
        env : Scope
            The local and global environment in which to evaluate an
            expression.

        Notes
        -----
        Must be implemented by subclasses.
        """


class NumExprEngine(AbstractEngine):
    """NumExpr engine class"""

    has_neg_frac = True

    def _evaluate(self):
        import numexpr as ne

        # convert the expression to a valid numexpr expression
        s = self.convert()

        env = self.expr.env
        scope = env.full_scope
        _check_ne_builtin_clash(self.expr)
        return ne.evaluate(s, local_dict=scope)


class PythonEngine(AbstractEngine):
    """
    Evaluate an expression in Python space.

    Mostly for testing purposes.
    """

    has_neg_frac = False

    def evaluate(self):
        return self.expr()

    def _evaluate(self) -> None:
        pass


ENGINES: dict[str, type[AbstractEngine]] = {
    "numexpr": NumExprEngine,
    "python": PythonEngine,
}

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\io\excel\_xlrd.py ===
from __future__ import annotations

from datetime import time
import math
from typing import TYPE_CHECKING

import numpy as np

from pandas.compat._optional import import_optional_dependency
from pandas.util._decorators import doc

from pandas.core.shared_docs import _shared_docs

from pandas.io.excel._base import BaseExcelReader

if TYPE_CHECKING:
    from xlrd import Book

    from pandas._typing import (
        Scalar,
        StorageOptions,
    )


class XlrdReader(BaseExcelReader["Book"]):
    @doc(storage_options=_shared_docs["storage_options"])
    def __init__(
        self,
        filepath_or_buffer,
        storage_options: StorageOptions | None = None,
        engine_kwargs: dict | None = None,
    ) -> None:
        """
        Reader using xlrd engine.

        Parameters
        ----------
        filepath_or_buffer : str, path object or Workbook
            Object to be parsed.
        {storage_options}
        engine_kwargs : dict, optional
            Arbitrary keyword arguments passed to excel engine.
        """
        err_msg = "Install xlrd >= 2.0.1 for xls Excel support"
        import_optional_dependency("xlrd", extra=err_msg)
        super().__init__(
            filepath_or_buffer,
            storage_options=storage_options,
            engine_kwargs=engine_kwargs,
        )

    @property
    def _workbook_class(self) -> type[Book]:
        from xlrd import Book

        return Book

    def load_workbook(self, filepath_or_buffer, engine_kwargs) -> Book:
        from xlrd import open_workbook

        if hasattr(filepath_or_buffer, "read"):
            data = filepath_or_buffer.read()
            return open_workbook(file_contents=data, **engine_kwargs)
        else:
            return open_workbook(filepath_or_buffer, **engine_kwargs)

    @property
    def sheet_names(self):
        return self.book.sheet_names()

    def get_sheet_by_name(self, name):
        self.raise_if_bad_sheet_by_name(name)
        return self.book.sheet_by_name(name)

    def get_sheet_by_index(self, index):
        self.raise_if_bad_sheet_by_index(index)
        return self.book.sheet_by_index(index)

    def get_sheet_data(
        self, sheet, file_rows_needed: int | None = None
    ) -> list[list[Scalar]]:
        from xlrd import (
            XL_CELL_BOOLEAN,
            XL_CELL_DATE,
            XL_CELL_ERROR,
            XL_CELL_NUMBER,
            xldate,
        )

        epoch1904 = self.book.datemode

        def _parse_cell(cell_contents, cell_typ):
            """
            converts the contents of the cell into a pandas appropriate object
            """
            if cell_typ == XL_CELL_DATE:
                # Use the newer xlrd datetime handling.
                try:
                    cell_contents = xldate.xldate_as_datetime(cell_contents, epoch1904)
                except OverflowError:
                    return cell_contents

                # Excel doesn't distinguish between dates and time,
                # so we treat dates on the epoch as times only.
                # Also, Excel supports 1900 and 1904 epochs.
                year = (cell_contents.timetuple())[0:3]
                if (not epoch1904 and year == (1899, 12, 31)) or (
                    epoch1904 and year == (1904, 1, 1)
                ):
                    cell_contents = time(
                        cell_contents.hour,
                        cell_contents.minute,
                        cell_contents.second,
                        cell_contents.microsecond,
                    )

            elif cell_typ == XL_CELL_ERROR:
                cell_contents = np.nan
            elif cell_typ == XL_CELL_BOOLEAN:
                cell_contents = bool(cell_contents)
            elif cell_typ == XL_CELL_NUMBER:
                # GH5394 - Excel 'numbers' are always floats
                # it's a minimal perf hit and less surprising
                if math.isfinite(cell_contents):
                    # GH54564 - don't attempt to convert NaN/Inf
                    val = int(cell_contents)
                    if val == cell_contents:
                        cell_contents = val
            return cell_contents

        data = []

        nrows = sheet.nrows
        if file_rows_needed is not None:
            nrows = min(nrows, file_rows_needed)
        for i in range(nrows):
            row = [
                _parse_cell(value, typ)
                for value, typ in zip(sheet.row_values(i), sheet.row_types(i))
            ]
            data.append(row)

        return data

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\mysql\expression.py ===
# dialects/mysql/expression.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors


from ... import exc
from ... import util
from ...sql import coercions
from ...sql import elements
from ...sql import operators
from ...sql import roles
from ...sql.base import _generative
from ...sql.base import Generative
from ...util.typing import Self


class match(Generative, elements.BinaryExpression):
    """Produce a ``MATCH (X, Y) AGAINST ('TEXT')`` clause.

    E.g.::

        from sqlalchemy import desc
        from sqlalchemy.dialects.mysql import match

        match_expr = match(
            users_table.c.firstname,
            users_table.c.lastname,
            against="Firstname Lastname",
        )

        stmt = (
            select(users_table)
            .where(match_expr.in_boolean_mode())
            .order_by(desc(match_expr))
        )

    Would produce SQL resembling:

    .. sourcecode:: sql

        SELECT id, firstname, lastname
        FROM user
        WHERE MATCH(firstname, lastname) AGAINST (:param_1 IN BOOLEAN MODE)
        ORDER BY MATCH(firstname, lastname) AGAINST (:param_2) DESC

    The :func:`_mysql.match` function is a standalone version of the
    :meth:`_sql.ColumnElement.match` method available on all
    SQL expressions, as when :meth:`_expression.ColumnElement.match` is
    used, but allows to pass multiple columns

    :param cols: column expressions to match against

    :param against: expression to be compared towards

    :param in_boolean_mode: boolean, set "boolean mode" to true

    :param in_natural_language_mode: boolean , set "natural language" to true

    :param with_query_expansion: boolean, set "query expansion" to true

    .. versionadded:: 1.4.19

    .. seealso::

        :meth:`_expression.ColumnElement.match`

    """

    __visit_name__ = "mysql_match"

    inherit_cache = True

    def __init__(self, *cols, **kw):
        if not cols:
            raise exc.ArgumentError("columns are required")

        against = kw.pop("against", None)

        if against is None:
            raise exc.ArgumentError("against is required")
        against = coercions.expect(
            roles.ExpressionElementRole,
            against,
        )

        left = elements.BooleanClauseList._construct_raw(
            operators.comma_op,
            clauses=cols,
        )
        left.group = False

        flags = util.immutabledict(
            {
                "mysql_boolean_mode": kw.pop("in_boolean_mode", False),
                "mysql_natural_language": kw.pop(
                    "in_natural_language_mode", False
                ),
                "mysql_query_expansion": kw.pop("with_query_expansion", False),
            }
        )

        if kw:
            raise exc.ArgumentError("unknown arguments: %s" % (", ".join(kw)))

        super().__init__(left, against, operators.match_op, modifiers=flags)

    @_generative
    def in_boolean_mode(self) -> Self:
        """Apply the "IN BOOLEAN MODE" modifier to the MATCH expression.

        :return: a new :class:`_mysql.match` instance with modifications
         applied.
        """

        self.modifiers = self.modifiers.union({"mysql_boolean_mode": True})
        return self

    @_generative
    def in_natural_language_mode(self) -> Self:
        """Apply the "IN NATURAL LANGUAGE MODE" modifier to the MATCH
        expression.

        :return: a new :class:`_mysql.match` instance with modifications
         applied.
        """

        self.modifiers = self.modifiers.union({"mysql_natural_language": True})
        return self

    @_generative
    def with_query_expansion(self) -> Self:
        """Apply the "WITH QUERY EXPANSION" modifier to the MATCH expression.

        :return: a new :class:`_mysql.match` instance with modifications
         applied.
        """

        self.modifiers = self.modifiers.union({"mysql_query_expansion": True})
        return self

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\diffgeom\rn.py ===
"""Predefined R^n manifolds together with common coord. systems.

Coordinate systems are predefined as well as the transformation laws between
them.

Coordinate functions can be accessed as attributes of the manifold (eg `R2.x`),
as attributes of the coordinate systems (eg `R2_r.x` and `R2_p.theta`), or by
using the usual `coord_sys.coord_function(index, name)` interface.
"""

from typing import Any
import warnings

from sympy.core.symbol import (Dummy, symbols)
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.trigonometric import (acos, atan2, cos, sin)
from .diffgeom import Manifold, Patch, CoordSystem

__all__ = [
    'R2', 'R2_origin', 'relations_2d', 'R2_r', 'R2_p',
    'R3', 'R3_origin', 'relations_3d', 'R3_r', 'R3_c', 'R3_s'
]

###############################################################################
# R2
###############################################################################
R2: Any = Manifold('R^2', 2)

R2_origin: Any = Patch('origin', R2)

x, y = symbols('x y', real=True)
r, theta = symbols('rho theta', nonnegative=True)

relations_2d = {
    ('rectangular', 'polar'): [(x, y), (sqrt(x**2 + y**2), atan2(y, x))],
    ('polar', 'rectangular'): [(r, theta), (r*cos(theta), r*sin(theta))],
}

R2_r: Any = CoordSystem('rectangular', R2_origin, (x, y), relations_2d)
R2_p: Any = CoordSystem('polar', R2_origin, (r, theta), relations_2d)

# support deprecated feature
with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    x, y, r, theta = symbols('x y r theta', cls=Dummy)
    R2_r.connect_to(R2_p, [x, y],
                        [sqrt(x**2 + y**2), atan2(y, x)],
                    inverse=False, fill_in_gaps=False)
    R2_p.connect_to(R2_r, [r, theta],
                        [r*cos(theta), r*sin(theta)],
                    inverse=False, fill_in_gaps=False)

# Defining the basis coordinate functions and adding shortcuts for them to the
# manifold and the patch.
R2.x, R2.y = R2_origin.x, R2_origin.y = R2_r.x, R2_r.y = R2_r.coord_functions()
R2.r, R2.theta = R2_origin.r, R2_origin.theta = R2_p.r, R2_p.theta = R2_p.coord_functions()

# Defining the basis vector fields and adding shortcuts for them to the
# manifold and the patch.
R2.e_x, R2.e_y = R2_origin.e_x, R2_origin.e_y = R2_r.e_x, R2_r.e_y = R2_r.base_vectors()
R2.e_r, R2.e_theta = R2_origin.e_r, R2_origin.e_theta = R2_p.e_r, R2_p.e_theta = R2_p.base_vectors()

# Defining the basis oneform fields and adding shortcuts for them to the
# manifold and the patch.
R2.dx, R2.dy = R2_origin.dx, R2_origin.dy = R2_r.dx, R2_r.dy = R2_r.base_oneforms()
R2.dr, R2.dtheta = R2_origin.dr, R2_origin.dtheta = R2_p.dr, R2_p.dtheta = R2_p.base_oneforms()

###############################################################################
# R3
###############################################################################
R3: Any = Manifold('R^3', 3)

R3_origin: Any = Patch('origin', R3)

x, y, z = symbols('x y z', real=True)
rho, psi, r, theta, phi = symbols('rho psi r theta phi', nonnegative=True)

relations_3d = {
    ('rectangular', 'cylindrical'): [(x, y, z),
                                     (sqrt(x**2 + y**2), atan2(y, x), z)],
    ('cylindrical', 'rectangular'): [(rho, psi, z),
                                     (rho*cos(psi), rho*sin(psi), z)],
    ('rectangular', 'spherical'): [(x, y, z),
                                   (sqrt(x**2 + y**2 + z**2),
                                    acos(z/sqrt(x**2 + y**2 + z**2)),
                                    atan2(y, x))],
    ('spherical', 'rectangular'): [(r, theta, phi),
                                   (r*sin(theta)*cos(phi),
                                    r*sin(theta)*sin(phi),
                                    r*cos(theta))],
    ('cylindrical', 'spherical'): [(rho, psi, z),
                                   (sqrt(rho**2 + z**2),
                                    acos(z/sqrt(rho**2 + z**2)),
                                    psi)],
    ('spherical', 'cylindrical'): [(r, theta, phi),
                                   (r*sin(theta), phi, r*cos(theta))],
}

R3_r: Any = CoordSystem('rectangular', R3_origin, (x, y, z), relations_3d)
R3_c: Any = CoordSystem('cylindrical', R3_origin, (rho, psi, z), relations_3d)
R3_s: Any = CoordSystem('spherical', R3_origin, (r, theta, phi), relations_3d)

# support deprecated feature
with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    x, y, z, rho, psi, r, theta, phi = symbols('x y z rho psi r theta phi', cls=Dummy)
    R3_r.connect_to(R3_c, [x, y, z],
                        [sqrt(x**2 + y**2), atan2(y, x), z],
                    inverse=False, fill_in_gaps=False)
    R3_c.connect_to(R3_r, [rho, psi, z],
                        [rho*cos(psi), rho*sin(psi), z],
                    inverse=False, fill_in_gaps=False)
    ## rectangular <-> spherical
    R3_r.connect_to(R3_s, [x, y, z],
                        [sqrt(x**2 + y**2 + z**2), acos(z/
                                sqrt(x**2 + y**2 + z**2)), atan2(y, x)],
                    inverse=False, fill_in_gaps=False)
    R3_s.connect_to(R3_r, [r, theta, phi],
                        [r*sin(theta)*cos(phi), r*sin(
                            theta)*sin(phi), r*cos(theta)],
                    inverse=False, fill_in_gaps=False)
    ## cylindrical <-> spherical
    R3_c.connect_to(R3_s, [rho, psi, z],
                        [sqrt(rho**2 + z**2), acos(z/sqrt(rho**2 + z**2)), psi],
                    inverse=False, fill_in_gaps=False)
    R3_s.connect_to(R3_c, [r, theta, phi],
                        [r*sin(theta), phi, r*cos(theta)],
                    inverse=False, fill_in_gaps=False)

# Defining the basis coordinate functions.
R3_r.x, R3_r.y, R3_r.z = R3_r.coord_functions()
R3_c.rho, R3_c.psi, R3_c.z = R3_c.coord_functions()
R3_s.r, R3_s.theta, R3_s.phi = R3_s.coord_functions()

# Defining the basis vector fields.
R3_r.e_x, R3_r.e_y, R3_r.e_z = R3_r.base_vectors()
R3_c.e_rho, R3_c.e_psi, R3_c.e_z = R3_c.base_vectors()
R3_s.e_r, R3_s.e_theta, R3_s.e_phi = R3_s.base_vectors()

# Defining the basis oneform fields.
R3_r.dx, R3_r.dy, R3_r.dz = R3_r.base_oneforms()
R3_c.drho, R3_c.dpsi, R3_c.dz = R3_c.base_oneforms()
R3_s.dr, R3_s.dtheta, R3_s.dphi = R3_s.base_oneforms()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\google\protobuf\internal\api_implementation.py ===
# Protocol Buffers - Google's data interchange format
# Copyright 2008 Google Inc.  All rights reserved.
#
# Use of this source code is governed by a BSD-style
# license that can be found in the LICENSE file or at
# https://developers.google.com/open-source/licenses/bsd

"""Determine which implementation of the protobuf API is used in this process.
"""

import importlib
import os
import sys
import warnings

_GOOGLE3_PYTHON_UPB_DEFAULT = True


def _ApiVersionToImplementationType(api_version):
  if api_version == 2:
    return 'cpp'
  if api_version == 1:
    raise ValueError('api_version=1 is no longer supported.')
  if api_version == 0:
    return 'python'
  return None


_implementation_type = None
try:
  # pylint: disable=g-import-not-at-top
  from google.protobuf.internal import _api_implementation
  # The compile-time constants in the _api_implementation module can be used to
  # switch to a certain implementation of the Python API at build time.
  _implementation_type = _ApiVersionToImplementationType(
      _api_implementation.api_version)
except ImportError:
  pass  # Unspecified by compiler flags.


def _CanImport(mod_name):
  try:
    mod = importlib.import_module(mod_name)
    # Work around a known issue in the classic bootstrap .par import hook.
    if not mod:
      raise ImportError(mod_name + ' import succeeded but was None')
    return True
  except ImportError:
    return False


if _implementation_type is None:
  if _CanImport('google._upb._message'):
    _implementation_type = 'upb'
  elif _CanImport('google.protobuf.pyext._message'):
    _implementation_type = 'cpp'
  else:
    _implementation_type = 'python'


# This environment variable can be used to switch to a certain implementation
# of the Python API, overriding the compile-time constants in the
# _api_implementation module. Right now only 'python', 'cpp' and 'upb' are
# valid values. Any other value will raise error.
_implementation_type = os.getenv('PROTOCOL_BUFFERS_PYTHON_IMPLEMENTATION',
                                 _implementation_type)

if _implementation_type not in ('python', 'cpp', 'upb'):
  raise ValueError('PROTOCOL_BUFFERS_PYTHON_IMPLEMENTATION {0} is not '
                   'supported. Please set to \'python\', \'cpp\' or '
                   '\'upb\'.'.format(_implementation_type))

if 'PyPy' in sys.version and _implementation_type == 'cpp':
  warnings.warn('PyPy does not work yet with cpp protocol buffers. '
                'Falling back to the python implementation.')
  _implementation_type = 'python'

_c_module = None

if _implementation_type == 'cpp':
  try:
    # pylint: disable=g-import-not-at-top
    from google.protobuf.pyext import _message
    sys.modules['google3.net.proto2.python.internal.cpp._message'] = _message
    _c_module = _message
    del _message
  except ImportError:
    # TODO: fail back to python
    warnings.warn(
        'Selected implementation cpp is not available.')
    pass

if _implementation_type == 'upb':
  try:
    # pylint: disable=g-import-not-at-top
    from google._upb import _message
    _c_module = _message
    del _message
  except ImportError:
    warnings.warn('Selected implementation upb is not available. '
                  'Falling back to the python implementation.')
    _implementation_type = 'python'
    pass

# Detect if serialization should be deterministic by default
try:
  # The presence of this module in a build allows the proto implementation to
  # be upgraded merely via build deps.
  #
  # NOTE: Merely importing this automatically enables deterministic proto
  # serialization for C++ code, but we still need to export it as a boolean so
  # that we can do the same for `_implementation_type == 'python'`.
  #
  # NOTE2: It is possible for C++ code to enable deterministic serialization by
  # default _without_ affecting Python code, if the C++ implementation is not in
  # use by this module.  That is intended behavior, so we don't actually expose
  # this boolean outside of this module.
  #
  # pylint: disable=g-import-not-at-top,unused-import
  from google.protobuf import enable_deterministic_proto_serialization
  _python_deterministic_proto_serialization = True
except ImportError:
  _python_deterministic_proto_serialization = False


# Usage of this function is discouraged. Clients shouldn't care which
# implementation of the API is in use. Note that there is no guarantee
# that differences between APIs will be maintained.
# Please don't use this function if possible.
def Type():
  return _implementation_type


# See comment on 'Type' above.
# TODO: Remove the API, it returns a constant. b/228102101
def Version():
  return 2


# For internal use only
def IsPythonDefaultSerializationDeterministic():
  return _python_deterministic_proto_serialization

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\plotting\_matplotlib\groupby.py ===
from __future__ import annotations

from typing import TYPE_CHECKING

import numpy as np

from pandas.core.dtypes.missing import remove_na_arraylike

from pandas import (
    MultiIndex,
    concat,
)

from pandas.plotting._matplotlib.misc import unpack_single_str_list

if TYPE_CHECKING:
    from collections.abc import Hashable

    from pandas._typing import IndexLabel

    from pandas import (
        DataFrame,
        Series,
    )


def create_iter_data_given_by(
    data: DataFrame, kind: str = "hist"
) -> dict[Hashable, DataFrame | Series]:
    """
    Create data for iteration given `by` is assigned or not, and it is only
    used in both hist and boxplot.

    If `by` is assigned, return a dictionary of DataFrames in which the key of
    dictionary is the values in groups.
    If `by` is not assigned, return input as is, and this preserves current
    status of iter_data.

    Parameters
    ----------
    data : reformatted grouped data from `_compute_plot_data` method.
    kind : str, plot kind. This function is only used for `hist` and `box` plots.

    Returns
    -------
    iter_data : DataFrame or Dictionary of DataFrames

    Examples
    --------
    If `by` is assigned:

    >>> import numpy as np
    >>> tuples = [('h1', 'a'), ('h1', 'b'), ('h2', 'a'), ('h2', 'b')]
    >>> mi = pd.MultiIndex.from_tuples(tuples)
    >>> value = [[1, 3, np.nan, np.nan],
    ...          [3, 4, np.nan, np.nan], [np.nan, np.nan, 5, 6]]
    >>> data = pd.DataFrame(value, columns=mi)
    >>> create_iter_data_given_by(data)
    {'h1':     h1
         a    b
    0  1.0  3.0
    1  3.0  4.0
    2  NaN  NaN, 'h2':     h2
         a    b
    0  NaN  NaN
    1  NaN  NaN
    2  5.0  6.0}
    """

    # For `hist` plot, before transformation, the values in level 0 are values
    # in groups and subplot titles, and later used for column subselection and
    # iteration; For `box` plot, values in level 1 are column names to show,
    # and are used for iteration and as subplots titles.
    if kind == "hist":
        level = 0
    else:
        level = 1

    # Select sub-columns based on the value of level of MI, and if `by` is
    # assigned, data must be a MI DataFrame
    assert isinstance(data.columns, MultiIndex)
    return {
        col: data.loc[:, data.columns.get_level_values(level) == col]
        for col in data.columns.levels[level]
    }


def reconstruct_data_with_by(
    data: DataFrame, by: IndexLabel, cols: IndexLabel
) -> DataFrame:
    """
    Internal function to group data, and reassign multiindex column names onto the
    result in order to let grouped data be used in _compute_plot_data method.

    Parameters
    ----------
    data : Original DataFrame to plot
    by : grouped `by` parameter selected by users
    cols : columns of data set (excluding columns used in `by`)

    Returns
    -------
    Output is the reconstructed DataFrame with MultiIndex columns. The first level
    of MI is unique values of groups, and second level of MI is the columns
    selected by users.

    Examples
    --------
    >>> d = {'h': ['h1', 'h1', 'h2'], 'a': [1, 3, 5], 'b': [3, 4, 6]}
    >>> df = pd.DataFrame(d)
    >>> reconstruct_data_with_by(df, by='h', cols=['a', 'b'])
       h1      h2
       a     b     a     b
    0  1.0   3.0   NaN   NaN
    1  3.0   4.0   NaN   NaN
    2  NaN   NaN   5.0   6.0
    """
    by_modified = unpack_single_str_list(by)
    grouped = data.groupby(by_modified)

    data_list = []
    for key, group in grouped:
        # error: List item 1 has incompatible type "Union[Hashable,
        # Sequence[Hashable]]"; expected "Iterable[Hashable]"
        columns = MultiIndex.from_product([[key], cols])  # type: ignore[list-item]
        sub_group = group[cols]
        sub_group.columns = columns
        data_list.append(sub_group)

    data = concat(data_list, axis=1)
    return data


def reformat_hist_y_given_by(y: np.ndarray, by: IndexLabel | None) -> np.ndarray:
    """Internal function to reformat y given `by` is applied or not for hist plot.

    If by is None, input y is 1-d with NaN removed; and if by is not None, groupby
    will take place and input y is multi-dimensional array.
    """
    if by is not None and len(y.shape) > 1:
        return np.array([remove_na_arraylike(col) for col in y.T]).T
    return remove_na_arraylike(y)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\timeouts.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from typing import TypedDict

    class JSONTimeouts(TypedDict, total=False):
        implicit: int
        pageLoad: int
        script: int

else:
    from typing import Dict

    JSONTimeouts = Dict[str, int]


class _TimeoutsDescriptor:
    """Get or set the value of the attributes listed below.

    _implicit_wait _page_load _script

    This does not set the value on the remote end.
    """

    def __init__(self, name):
        self.name = name

    def __get__(self, obj, cls) -> float:
        return getattr(obj, self.name) / 1000

    def __set__(self, obj, value) -> None:
        converted_value = getattr(obj, "_convert")(value)
        setattr(obj, self.name, converted_value)


class Timeouts:
    def __init__(self, implicit_wait: float = 0, page_load: float = 0, script: float = 0) -> None:
        """Create a new Timeouts object.

        This implements https://w3c.github.io/webdriver/#timeouts.

        :Args:
         - implicit_wait - Either an int or a float. Set how many
            seconds to wait when searching for elements before
            throwing an error.
         - page_load - Either an int or a float. Set how many seconds
            to wait for a page load to complete before throwing
            an error.
         - script - Either an int or a float. Set how many seconds to
            wait for an asynchronous script to finish execution
            before throwing an error.
        """

        self._implicit_wait = self._convert(implicit_wait)
        self._page_load = self._convert(page_load)
        self._script = self._convert(script)

    # Creating descriptor objects
    implicit_wait = _TimeoutsDescriptor("_implicit_wait")
    """Get or set how many seconds to wait when searching for elements.

    This does not set the value on the remote end.

    Usage:
    ------
    - Get
        - `self.implicit_wait`
    - Set
        - `self.implicit_wait` = `value`

    Parameters:
    -----------
    `value`: `float`
    """

    page_load = _TimeoutsDescriptor("_page_load")
    """Get or set how many seconds to wait for the page to load.

    This does not set the value on the remote end.

    Usage:
    ------
    - Get
        - `self.page_load`
    - Set
        - `self.page_load` = `value`

    Parameters:
    -----------
    `value`: `float`
    """

    script = _TimeoutsDescriptor("_script")
    """Get or set how many seconds to wait for an asynchronous script to finish
    execution.

    This does not set the value on the remote end.

    Usage:
    ------
    - Get
        - `self.script`
    - Set
        - `self.script` = `value`

    Parameters:
    -----------
    `value`: `float`
    """

    def _convert(self, timeout: float) -> int:
        if isinstance(timeout, (int, float)):
            return int(float(timeout) * 1000)
        raise TypeError("Timeouts can only be an int or a float")

    def _to_json(self) -> JSONTimeouts:
        timeouts: JSONTimeouts = {}
        if self._implicit_wait:
            timeouts["implicit"] = self._implicit_wait
        if self._page_load:
            timeouts["pageLoad"] = self._page_load
        if self._script:
            timeouts["script"] = self._script

        return timeouts

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\quantization\preprocess.py ===
# --------------------------------------------------------------------------
# Copyright (c) Microsoft, Intel Corporation. All rights reserved.
# Licensed under the MIT License. See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------

import argparse
import logging
import sys

from .shape_inference import quant_pre_process

logger = logging.getLogger(__name__)


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="""Model optimizer and shape inferencer, in preparation for quantization,
Consists of three optional steps:
1. Symbolic shape inference (best for transformer models).
2. Model optimization.
3. ONNX shape inference.

Model quantization with QDQ format, i.e. inserting QuantizeLinear/DeQuantizeLinear on
the tensor, requires tensor shape information to perform its best. Currently, shape inferencing
works best with optimized model. As a result, it is highly recommended to run quantization
on optimized model with shape information. This is the tool for optimization and shape
inferencing.

Essentially this tool performs the following three (skippable) steps:

1. Symbolic shape inference.
2. Model optimization
3. ONNX shape inference"""
    )

    parser.add_argument("--input", required=True, help="Path to the input model file")
    parser.add_argument("--output", required=True, help="Path to the output model file")
    parser.add_argument(
        "--skip_optimization",
        type=bool,
        default=False,
        help="Skip model optimization step if true. It's a known issue that ORT"
        " optimization has difficulty with model size greater than 2GB, rerun with"
        " this option to get around this issue.",
    )
    parser.add_argument(
        "--skip_onnx_shape",
        type=bool,
        default=False,
        help="Skip ONNX shape inference. Symbolic shape inference is most effective"
        " with transformer based models. Skipping all shape inferences may"
        " reduce the effectiveness of quantization, as a tensor with unknown"
        " shape can not be quantized.",
    )
    parser.add_argument(
        "--skip_symbolic_shape",
        type=bool,
        default=False,
        help="Skip symbolic shape inference. Symbolic shape inference is most"
        " effective with transformer based models. Skipping all shape"
        " inferences may reduce the effectiveness of quantization, as a tensor"
        " with unknown shape can not be quantized.",
    )
    parser.add_argument(
        "--auto_merge",
        help="Automatically merge symbolic dims when confliction happens",
        action="store_true",
        default=False,
    )
    parser.add_argument(
        "--int_max",
        help="maximum value for integer to be treated as boundless for ops like slice",
        type=int,
        default=2**31 - 1,
    )
    parser.add_argument(
        "--guess_output_rank",
        help="guess output rank to be the same as input 0 for unknown ops",
        action="store_true",
        default=False,
    )
    parser.add_argument(
        "--verbose",
        help="Prints detailed logs of inference, 0: turn off, 1: warnings, 3: detailed",
        type=int,
        default=0,
    )
    parser.add_argument(
        "--save_as_external_data",
        help="Saving an ONNX model to external data",
        action="store_true",
        default=False,
    )
    parser.add_argument(
        "--all_tensors_to_one_file",
        help="Saving all the external data to one file",
        action="store_true",
        default=False,
    )
    parser.add_argument(
        "--external_data_location",
        help="The file location to save the external file",
        default=None,
    )
    parser.add_argument(
        "--external_data_size_threshold",
        help="The size threshold for external data",
        type=int,
        default=1024,
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_arguments()
    if args.skip_optimization and args.skip_onnx_shape and args.skip_symbolic_shape:
        logger.error("Skipping all three steps, nothing to be done. Quitting...")
        sys.exit()

    if (not args.skip_optimization) and args.save_as_external_data:
        logger.error("ORT model optimization does not support external data yet!")
        sys.exit()

    logger.info("input model: %s", args.input)
    logger.info("output model: %s", args.output)
    quant_pre_process(
        args.input,
        args.output,
        args.skip_optimization,
        args.skip_onnx_shape,
        args.skip_symbolic_shape,
        args.auto_merge,
        args.int_max,
        args.guess_output_rank,
        args.verbose,
        args.save_as_external_data,
        args.all_tensors_to_one_file,
        args.external_data_location,
        args.external_data_size_threshold,
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\tools\ort_format_model\ort_flatbuffers_py\fbs\ModuleState.py ===
# automatically generated by the FlatBuffers compiler, do not modify

# namespace: fbs

import flatbuffers
from flatbuffers.compat import import_numpy
np = import_numpy()

class ModuleState(object):
    __slots__ = ['_tab']

    @classmethod
    def GetRootAs(cls, buf, offset=0):
        n = flatbuffers.encode.Get(flatbuffers.packer.uoffset, buf, offset)
        x = ModuleState()
        x.Init(buf, n + offset)
        return x

    @classmethod
    def GetRootAsModuleState(cls, buf, offset=0):
        """This method is deprecated. Please switch to GetRootAs."""
        return cls.GetRootAs(buf, offset)
    @classmethod
    def ModuleStateBufferHasIdentifier(cls, buf, offset, size_prefixed=False):
        return flatbuffers.util.BufferHasIdentifier(buf, offset, b"\x4F\x44\x54\x43", size_prefixed=size_prefixed)

    # ModuleState
    def Init(self, buf, pos):
        self._tab = flatbuffers.table.Table(buf, pos)

    # ModuleState
    def RequiresGradParams(self, j):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(4))
        if o != 0:
            x = self._tab.Vector(o)
            x += flatbuffers.number_types.UOffsetTFlags.py_type(j) * 4
            x = self._tab.Indirect(x)
            from ort_flatbuffers_py.fbs.Tensor import Tensor
            obj = Tensor()
            obj.Init(self._tab.Bytes, x)
            return obj
        return None

    # ModuleState
    def RequiresGradParamsLength(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(4))
        if o != 0:
            return self._tab.VectorLen(o)
        return 0

    # ModuleState
    def RequiresGradParamsIsNone(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(4))
        return o == 0

    # ModuleState
    def FrozenParams(self, j):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        if o != 0:
            x = self._tab.Vector(o)
            x += flatbuffers.number_types.UOffsetTFlags.py_type(j) * 4
            x = self._tab.Indirect(x)
            from ort_flatbuffers_py.fbs.Tensor import Tensor
            obj = Tensor()
            obj.Init(self._tab.Bytes, x)
            return obj
        return None

    # ModuleState
    def FrozenParamsLength(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        if o != 0:
            return self._tab.VectorLen(o)
        return 0

    # ModuleState
    def FrozenParamsIsNone(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        return o == 0

    # ModuleState
    def IsNominalState(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(8))
        if o != 0:
            return bool(self._tab.Get(flatbuffers.number_types.BoolFlags, o + self._tab.Pos))
        return False

    # ModuleState
    def HasExternalData(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(10))
        if o != 0:
            return bool(self._tab.Get(flatbuffers.number_types.BoolFlags, o + self._tab.Pos))
        return False

def ModuleStateStart(builder):
    builder.StartObject(4)

def Start(builder):
    ModuleStateStart(builder)

def ModuleStateAddRequiresGradParams(builder, requiresGradParams):
    builder.PrependUOffsetTRelativeSlot(0, flatbuffers.number_types.UOffsetTFlags.py_type(requiresGradParams), 0)

def AddRequiresGradParams(builder, requiresGradParams):
    ModuleStateAddRequiresGradParams(builder, requiresGradParams)

def ModuleStateStartRequiresGradParamsVector(builder, numElems):
    return builder.StartVector(4, numElems, 4)

def StartRequiresGradParamsVector(builder, numElems: int) -> int:
    return ModuleStateStartRequiresGradParamsVector(builder, numElems)

def ModuleStateAddFrozenParams(builder, frozenParams):
    builder.PrependUOffsetTRelativeSlot(1, flatbuffers.number_types.UOffsetTFlags.py_type(frozenParams), 0)

def AddFrozenParams(builder, frozenParams):
    ModuleStateAddFrozenParams(builder, frozenParams)

def ModuleStateStartFrozenParamsVector(builder, numElems):
    return builder.StartVector(4, numElems, 4)

def StartFrozenParamsVector(builder, numElems: int) -> int:
    return ModuleStateStartFrozenParamsVector(builder, numElems)

def ModuleStateAddIsNominalState(builder, isNominalState):
    builder.PrependBoolSlot(2, isNominalState, 0)

def AddIsNominalState(builder, isNominalState):
    ModuleStateAddIsNominalState(builder, isNominalState)

def ModuleStateAddHasExternalData(builder, hasExternalData):
    builder.PrependBoolSlot(3, hasExternalData, 0)

def AddHasExternalData(builder, hasExternalData):
    ModuleStateAddHasExternalData(builder, hasExternalData)

def ModuleStateEnd(builder):
    return builder.EndObject()

def End(builder):
    return ModuleStateEnd(builder)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_base.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
from collections import defaultdict
from collections.abc import Sequence
from logging import getLogger
from typing import Any

import numpy as np
from onnx import NodeProto, TensorProto, helper
from onnx_model import OnnxModel

logger = getLogger(__name__)


class Fusion:
    """
    Base class for Graph Fusion
    """

    def __init__(
        self,
        model: OnnxModel,
        fused_op_type: str,
        search_op_types: str | list[str],
        description: str = "",
    ):
        self.search_op_types: list[str] = [search_op_types] if isinstance(search_op_types, str) else search_op_types
        self.fused_op_type: str = fused_op_type
        self.description: str = f"{fused_op_type}({description})" if description else fused_op_type
        self.model: OnnxModel = model
        self.nodes_to_remove: list = []
        self.nodes_to_add: list = []
        self.prune_graph: bool = False
        self.node_name_to_graph_name: dict = {}
        self.this_graph_name: str | None = None
        # It is optional that subclass updates fused_count since we will also check nodes_to_add to get counter.
        self.fused_count: defaultdict = defaultdict(int)

    def increase_counter(self, fused_op_name: str):
        """
        Increase counter of a fused operator.
        """
        self.fused_count[fused_op_name] += 1

    def fuse(
        self,
        node: NodeProto,
        input_name_to_nodes: dict[str, list[NodeProto]],
        output_name_to_node: dict[str, NodeProto],
    ):
        """Interface for fusion that starts from a node"""
        raise NotImplementedError

    def apply(self):
        """
        Apply graph fusion on the whole model graph.
        It searched nodes of given operators, and start fusion on each of those nodes.
        """
        logger.debug(f"start {self.description} fusion...")
        input_name_to_nodes = self.model.input_name_to_nodes()
        output_name_to_node = self.model.output_name_to_node()

        # This assumes that two search ops will not be fused at same time!
        for search_op_type in self.search_op_types:
            for node in self.model.get_nodes_by_op_type(search_op_type):
                graph = self.model.get_graph_by_node(node)
                if graph is None:
                    raise Exception("Can not find node in any graph")
                self.this_graph_name = graph.name
                self.fuse(node, input_name_to_nodes, output_name_to_node)

        op_list = [node.op_type for node in self.nodes_to_add]
        if self.fused_count:
            for key, value in self.fused_count.items():
                if value:
                    logger.info(f"Fused {key}: {value}")
        else:
            count = op_list.count(self.fused_op_type)
            if count > 0:
                logger.info(f"Fused {self.description}: {count}")

        self.model.remove_nodes(self.nodes_to_remove)
        self.model.add_nodes(self.nodes_to_add, self.node_name_to_graph_name)

        if self.prune_graph:
            self.model.prune_graph()
        elif self.nodes_to_remove or self.nodes_to_add:
            self.model.update_graph()

    def add_initializer(self, name: str, data_type: int, dims: Sequence[int], vals: Any, raw: bool = True):
        if raw:
            np_type = helper.tensor_dtype_to_np_dtype(data_type)
            if not isinstance(vals, np.ndarray):
                bytes = np.array(vals, dtype=np_type).tobytes()
            else:
                bytes = vals.astype(np_type).tobytes()
            tensor = helper.make_tensor(
                name=name,
                data_type=data_type,
                dims=dims,
                vals=bytes,
                raw=True,
            )
        else:
            tensor = helper.make_tensor(
                name=name,
                data_type=data_type,
                dims=dims,
                vals=vals,
                raw=False,
            )

        self.model.add_initializer(tensor, self.this_graph_name)
        return tensor

    def remove_initializer(self, tensor: TensorProto):
        self.model.remove_initializer(tensor)

    def add_nodes_to_remove(self, nodes: list[NodeProto]):
        # Some nodes are shared between paths (e.g. rotary embedding nodes in the Q and K paths).
        # When path A is fused, its shared nodes are added to `self.nodes_to_remove`. But when path B
        # is fused, its shared nodes are also added to `self.nodes_to_remove`. When the nodes are
        # iteratively removed from `self.nodes_to_remove`, path A's shared nodes are removed first.
        # Since path A's shared nodes are removed, path B's shared nodes are not removed because they
        # were previously removed for path A. This causes an error to print in remove_node that a node
        # has failed to be removed.
        #
        # To avoid this error, we pre-emptively check if the shared nodes are already in `self.nodes_to_remove`.
        # We could alternatively convert `self.nodes_to_remove` to a set to avoid this issue, but there could
        # be scenarios where the nodes need to be removed in a specific order and converting to a set would
        # lose this order.
        for node in nodes:
            if node not in self.nodes_to_remove:
                self.nodes_to_remove.append(node)

    def add_nodes_to_remove_with_nodes_to_keep(self, nodes: list[NodeProto], nodes_to_keep: list[NodeProto]):
        for node in nodes:
            if node not in self.nodes_to_remove and node not in nodes_to_keep:
                self.nodes_to_remove.append(node)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\onnx_model_bart.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
import logging

from fusion_attention import AttentionMask
from fusion_bart_attention import FusionBartAttention
from fusion_options import FusionOptions
from fusion_reshape import FusionReshape
from onnx import numpy_helper
from onnx_model import OnnxModel
from onnx_model_bert import BertOnnxModel

logger = logging.getLogger(__name__)


class FusionBartReshape(FusionReshape):
    def __init__(self, model: OnnxModel):
        super().__init__(model)

    def fuse(self, reshape_node, input_name_to_nodes, output_name_to_node):
        if reshape_node.input[1] not in output_name_to_node:
            return

        concat_node = output_name_to_node[reshape_node.input[1]]
        if concat_node.op_type != "Concat" or len(concat_node.input) != 4:
            return

        path0 = self.model.match_parent_path(
            concat_node,
            ["Unsqueeze", "Gather", "Shape"],
            [0, 0, 0],
            output_name_to_node,
        )
        if path0 is None:
            return

        (_, gather_0, shape_0) = path0

        shape = []
        gather_value = self.model.get_constant_value(gather_0.input[1])
        if gather_value == 0:
            shape.append(0)

        path1 = self.model.match_parent_path(
            concat_node,
            ["Unsqueeze", "Gather", "Shape"],
            [1, 0, 0],
            output_name_to_node,
        )
        if path1 is None:
            input_1_proto = self.model.get_initializer(concat_node.input[1])
            input_2_proto = self.model.get_initializer(concat_node.input[2])
            input_3_proto = self.model.get_initializer(concat_node.input[3])
            if input_1_proto is None or input_2_proto is None or input_3_proto is None:
                return

            input_1 = numpy_helper.to_array(input_1_proto)
            input_2 = numpy_helper.to_array(input_2_proto)
            input_3 = numpy_helper.to_array(input_3_proto)
            if len(input_1) != 1 or len(input_2) != 1 or len(input_3) != 1:
                return

            if not (input_1[0] == -1 and input_2[0] > 0 and input_3[0] > 0):
                return

            shape.extend(input_1)
            shape.extend(input_2)
            shape.extend(input_3)
            gemm_path_with_bias = self.model.match_parent_path(
                reshape_node, ["Add", "MatMul"], [0, 1], output_name_to_node
            )
            gemm_path_no_bias = self.model.match_parent_path(reshape_node, ["MatMul"], [0], output_name_to_node)
            if gemm_path_with_bias is not None:
                gemm_path = gemm_path_with_bias
            elif gemm_path_no_bias is not None:
                gemm_path = gemm_path_no_bias
            else:
                return

            top_matmul = gemm_path[-1]
            root_input = top_matmul.input[0]

            self.replace_reshape_node(shape, reshape_node, concat_node)
        else:
            (_, gather_1, shape_1) = path1

            gather_value = self.model.get_constant_value(gather_1.input[1])
            if gather_value == 1:
                shape.append(0)

            input_2_proto = self.model.get_initializer(concat_node.input[2])
            input_3_proto = self.model.get_initializer(concat_node.input[3])
            if input_2_proto is None or input_3_proto is None:
                return

            input_2 = numpy_helper.to_array(input_2_proto)
            input_3 = numpy_helper.to_array(input_3_proto)
            if len(input_2) != 1 or len(input_3) != 1:
                return

            if not (input_2[0] > 0 and input_3[0] > 0):
                return

            shape.extend(input_2)
            shape.extend(input_3)
            gemm_path = self.model.match_parent_path(
                reshape_node, ["Mul", "Add", "MatMul"], [0, 0, 1], output_name_to_node
            )
            if gemm_path is None:
                return

            top_matmul = gemm_path[-1]
            root_input = top_matmul.input[0]
            if shape_0.input[0] != root_input or shape_1.input[0] != root_input:
                return

            self.replace_reshape_node(shape, reshape_node, concat_node)


class BartOnnxModel(BertOnnxModel):
    def __init__(self, model, num_heads, hidden_size, model_impl="hf"):
        super().__init__(model, num_heads, hidden_size)
        self.attention_mask = AttentionMask(self)
        self.attention_fusion = FusionBartAttention(self, self.hidden_size, self.num_heads, self.attention_mask)
        self.bart_reshape_fusion_preprocess = FusionBartReshape(self)

    def optimize(self, options: FusionOptions | None = None, add_dynamic_axes: bool = False):
        self.attention_fusion.use_multi_head_attention = False if options is None else options.use_multi_head_attention
        self.attention_fusion.disable_multi_head_attention_bias = (
            False if options is None else options.disable_multi_head_attention_bias
        )
        super().optimize(options, add_dynamic_axes)

    def fuse_attention(self):
        self.attention_fusion.apply()

    def preprocess(self):
        self.adjust_reshape_and_expand()
        self.bart_reshape_fusion_preprocess.apply()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\worksheet\merge.py ===
# Copyright (c) 2010-2024 openpyxl

import copy

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Integer,
    Sequence,
)

from openpyxl.cell.cell import MergedCell
from openpyxl.styles.borders import Border

from .cell_range import CellRange


class MergeCell(CellRange):

    tagname = "mergeCell"
    ref = CellRange.coord

    __attrs__ = ("ref",)


    def __init__(self,
                 ref=None,
                ):
        super().__init__(ref)


    def __copy__(self):
        return self.__class__(self.ref)


class MergeCells(Serialisable):

    tagname = "mergeCells"

    count = Integer(allow_none=True)
    mergeCell = Sequence(expected_type=MergeCell, )

    __elements__ = ('mergeCell',)
    __attrs__ = ('count',)

    def __init__(self,
                 count=None,
                 mergeCell=(),
                ):
        self.mergeCell = mergeCell


    @property
    def count(self):
        return len(self.mergeCell)


class MergedCellRange(CellRange):

    """
    MergedCellRange stores the border information of a merged cell in the top
    left cell of the merged cell.
    The remaining cells in the merged cell are stored as MergedCell objects and
    get their border information from the upper left cell.
    """

    def __init__(self, worksheet, coord):
        self.ws = worksheet
        super().__init__(range_string=coord)
        self.start_cell = None
        self._get_borders()


    def _get_borders(self):
        """
        If the upper left cell of the merged cell does not yet exist, it is
        created.
        The upper left cell gets the border information of the bottom and right
        border from the bottom right cell of the merged cell, if available.
        """

        # Top-left cell.
        self.start_cell = self.ws._cells.get((self.min_row, self.min_col))
        if self.start_cell is None:
            self.start_cell = self.ws.cell(row=self.min_row, column=self.min_col)

        # Bottom-right cell
        end_cell = self.ws._cells.get((self.max_row, self.max_col))
        if end_cell is not None:
            self.start_cell.border += Border(right=end_cell.border.right,
                                             bottom=end_cell.border.bottom)


    def format(self):
        """
        Each cell of the merged cell is created as MergedCell if it does not
        already exist.

        The MergedCells at the edge of the merged cell gets its borders from
        the upper left cell.

         - The top MergedCells get the top border from the top left cell.
         - The bottom MergedCells get the bottom border from the top left cell.
         - The left MergedCells get the left border from the top left cell.
         - The right MergedCells get the right border from the top left cell.
        """

        names = ['top', 'left', 'right', 'bottom']

        for name in names:
            side = getattr(self.start_cell.border, name)
            if side and side.style is None:
                continue # don't need to do anything if there is no border style
            border = Border(**{name:side})
            for coord in getattr(self, name):
                cell = self.ws._cells.get(coord)
                if cell is None:
                    row, col = coord
                    cell = MergedCell(self.ws, row=row, column=col)
                    self.ws._cells[(cell.row, cell.column)] = cell
                cell.border += border

        protected = self.start_cell.protection is not None
        if protected:
            protection = copy.copy(self.start_cell.protection)
        for coord in self.cells:
            cell = self.ws._cells.get(coord)
            if cell is None:
                row, col = coord
                cell = MergedCell(self.ws, row=row, column=col)
                self.ws._cells[(cell.row, cell.column)] = cell

            if protected:
                cell.protection = protection


    def __contains__(self, coord):
        return coord in CellRange(self.coord)


    def __copy__(self):
        return self.__class__(self.ws, self.coord)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\padding.py ===
from typing import TYPE_CHECKING, List, Optional, Tuple, Union

if TYPE_CHECKING:
    from .console import (
        Console,
        ConsoleOptions,
        RenderableType,
        RenderResult,
    )

from .jupyter import JupyterMixin
from .measure import Measurement
from .segment import Segment
from .style import Style

PaddingDimensions = Union[int, Tuple[int], Tuple[int, int], Tuple[int, int, int, int]]


class Padding(JupyterMixin):
    """Draw space around content.

    Example:
        >>> print(Padding("Hello", (2, 4), style="on blue"))

    Args:
        renderable (RenderableType): String or other renderable.
        pad (Union[int, Tuple[int]]): Padding for top, right, bottom, and left borders.
            May be specified with 1, 2, or 4 integers (CSS style).
        style (Union[str, Style], optional): Style for padding characters. Defaults to "none".
        expand (bool, optional): Expand padding to fit available width. Defaults to True.
    """

    def __init__(
        self,
        renderable: "RenderableType",
        pad: "PaddingDimensions" = (0, 0, 0, 0),
        *,
        style: Union[str, Style] = "none",
        expand: bool = True,
    ):
        self.renderable = renderable
        self.top, self.right, self.bottom, self.left = self.unpack(pad)
        self.style = style
        self.expand = expand

    @classmethod
    def indent(cls, renderable: "RenderableType", level: int) -> "Padding":
        """Make padding instance to render an indent.

        Args:
            renderable (RenderableType): String or other renderable.
            level (int): Number of characters to indent.

        Returns:
            Padding: A Padding instance.
        """

        return Padding(renderable, pad=(0, 0, 0, level), expand=False)

    @staticmethod
    def unpack(pad: "PaddingDimensions") -> Tuple[int, int, int, int]:
        """Unpack padding specified in CSS style."""
        if isinstance(pad, int):
            return (pad, pad, pad, pad)
        if len(pad) == 1:
            _pad = pad[0]
            return (_pad, _pad, _pad, _pad)
        if len(pad) == 2:
            pad_top, pad_right = pad
            return (pad_top, pad_right, pad_top, pad_right)
        if len(pad) == 4:
            top, right, bottom, left = pad
            return (top, right, bottom, left)
        raise ValueError(f"1, 2 or 4 integers required for padding; {len(pad)} given")

    def __repr__(self) -> str:
        return f"Padding({self.renderable!r}, ({self.top},{self.right},{self.bottom},{self.left}))"

    def __rich_console__(
        self, console: "Console", options: "ConsoleOptions"
    ) -> "RenderResult":
        style = console.get_style(self.style)
        if self.expand:
            width = options.max_width
        else:
            width = min(
                Measurement.get(console, options, self.renderable).maximum
                + self.left
                + self.right,
                options.max_width,
            )
        render_options = options.update_width(width - self.left - self.right)
        if render_options.height is not None:
            render_options = render_options.update_height(
                height=render_options.height - self.top - self.bottom
            )
        lines = console.render_lines(
            self.renderable, render_options, style=style, pad=True
        )
        _Segment = Segment

        left = _Segment(" " * self.left, style) if self.left else None
        right = (
            [_Segment(f'{" " * self.right}', style), _Segment.line()]
            if self.right
            else [_Segment.line()]
        )
        blank_line: Optional[List[Segment]] = None
        if self.top:
            blank_line = [_Segment(f'{" " * width}\n', style)]
            yield from blank_line * self.top
        if left:
            for line in lines:
                yield left
                yield from line
                yield from right
        else:
            for line in lines:
                yield from line
                yield from right
        if self.bottom:
            blank_line = blank_line or [_Segment(f'{" " * width}\n', style)]
            yield from blank_line * self.bottom

    def __rich_measure__(
        self, console: "Console", options: "ConsoleOptions"
    ) -> "Measurement":
        max_width = options.max_width
        extra_width = self.left + self.right
        if max_width - extra_width < 1:
            return Measurement(max_width, max_width)
        measure_min, measure_max = Measurement.get(console, options, self.renderable)
        measurement = Measurement(measure_min + extra_width, measure_max + extra_width)
        measurement = measurement.with_maximum(max_width)
        return measurement


if __name__ == "__main__":  #  pragma: no cover
    from pip._vendor.rich import print

    print(Padding("Hello, World", (2, 4), style="on blue"))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\type_tests\path.py ===
"""Path wrapping is quite complex, ensure all methods are understood as wrapped correctly."""

import io
import os
import pathlib
import sys
from typing import IO, Any, BinaryIO

import trio
from trio._file_io import AsyncIOWrapper
from typing_extensions import assert_type


def operator_checks(text: str, tpath: trio.Path, ppath: pathlib.Path) -> None:
    """Verify operators produce the right results."""
    assert_type(tpath / ppath, trio.Path)
    assert_type(tpath / tpath, trio.Path)
    assert_type(tpath / text, trio.Path)
    assert_type(text / tpath, trio.Path)

    assert_type(tpath > tpath, bool)
    assert_type(tpath >= tpath, bool)
    assert_type(tpath < tpath, bool)
    assert_type(tpath <= tpath, bool)

    assert_type(tpath > ppath, bool)
    assert_type(tpath >= ppath, bool)
    assert_type(tpath < ppath, bool)
    assert_type(tpath <= ppath, bool)

    assert_type(ppath > tpath, bool)
    assert_type(ppath >= tpath, bool)
    assert_type(ppath < tpath, bool)
    assert_type(ppath <= tpath, bool)


def sync_attrs(path: trio.Path) -> None:
    assert_type(path.parts, tuple[str, ...])
    assert_type(path.drive, str)
    assert_type(path.root, str)
    assert_type(path.anchor, str)
    assert_type(path.parents[3], trio.Path)
    assert_type(path.parent, trio.Path)
    assert_type(path.name, str)
    assert_type(path.suffix, str)
    assert_type(path.suffixes, list[str])
    assert_type(path.stem, str)
    assert_type(path.as_posix(), str)
    assert_type(path.as_uri(), str)
    assert_type(path.is_absolute(), bool)
    assert_type(path.is_relative_to(path), bool)
    assert_type(path.is_reserved(), bool)
    assert_type(path.joinpath(path, "folder"), trio.Path)
    assert_type(path.match("*.py"), bool)
    assert_type(path.relative_to("/usr"), trio.Path)
    if sys.version_info >= (3, 12):
        assert_type(path.relative_to("/", walk_up=True), trio.Path)
    assert_type(path.with_name("filename.txt"), trio.Path)
    assert_type(path.with_stem("readme"), trio.Path)
    assert_type(path.with_suffix(".log"), trio.Path)


async def async_attrs(path: trio.Path) -> None:
    assert_type(await trio.Path.cwd(), trio.Path)
    assert_type(await trio.Path.home(), trio.Path)
    assert_type(await path.stat(), os.stat_result)
    assert_type(await path.chmod(0o777), None)
    assert_type(await path.exists(), bool)
    assert_type(await path.expanduser(), trio.Path)
    for result in await path.glob("*.py"):
        assert_type(result, trio.Path)
    if sys.platform != "win32":
        assert_type(await path.group(), str)
    assert_type(await path.is_dir(), bool)
    assert_type(await path.is_file(), bool)
    if sys.version_info >= (3, 12):
        assert_type(await path.is_junction(), bool)
    if sys.platform != "win32":
        assert_type(await path.is_mount(), bool)
    assert_type(await path.is_symlink(), bool)
    assert_type(await path.is_socket(), bool)
    assert_type(await path.is_fifo(), bool)
    assert_type(await path.is_block_device(), bool)
    assert_type(await path.is_char_device(), bool)
    for child_iter in await path.iterdir():
        assert_type(child_iter, trio.Path)
    # TODO: Path.walk() in 3.12
    assert_type(await path.lchmod(0o111), None)
    assert_type(await path.lstat(), os.stat_result)
    assert_type(await path.mkdir(mode=0o777, parents=True, exist_ok=False), None)
    # Open done separately.
    if sys.platform != "win32":
        assert_type(await path.owner(), str)
    assert_type(await path.read_bytes(), bytes)
    assert_type(await path.read_text(encoding="utf16", errors="replace"), str)
    assert_type(await path.readlink(), trio.Path)
    assert_type(await path.rename("another"), trio.Path)
    assert_type(await path.replace(path), trio.Path)
    assert_type(await path.resolve(), trio.Path)
    for child_glob in await path.glob("*.py"):
        assert_type(child_glob, trio.Path)
    for child_rglob in await path.rglob("*.py"):
        assert_type(child_rglob, trio.Path)
    assert_type(await path.rmdir(), None)
    assert_type(await path.samefile("something_else"), bool)
    assert_type(await path.symlink_to("somewhere"), None)
    if sys.version_info >= (3, 10):
        assert_type(await path.hardlink_to("elsewhere"), None)
    assert_type(await path.touch(), None)
    assert_type(await path.unlink(missing_ok=True), None)
    assert_type(await path.write_bytes(b"123"), int)
    assert_type(
        await path.write_text("hello", encoding="utf32le", errors="ignore"),
        int,
    )


async def open_results(path: trio.Path, some_int: int, some_str: str) -> None:
    # Check the overloads.
    assert_type(await path.open(), AsyncIOWrapper[io.TextIOWrapper])
    assert_type(await path.open("r"), AsyncIOWrapper[io.TextIOWrapper])
    assert_type(await path.open("r+"), AsyncIOWrapper[io.TextIOWrapper])
    assert_type(await path.open("w"), AsyncIOWrapper[io.TextIOWrapper])
    assert_type(await path.open("rb", buffering=0), AsyncIOWrapper[io.FileIO])
    assert_type(await path.open("rb+"), AsyncIOWrapper[io.BufferedRandom])
    assert_type(await path.open("wb"), AsyncIOWrapper[io.BufferedWriter])
    assert_type(await path.open("rb"), AsyncIOWrapper[io.BufferedReader])
    assert_type(await path.open("rb", buffering=some_int), AsyncIOWrapper[BinaryIO])
    assert_type(await path.open(some_str), AsyncIOWrapper[IO[Any]])

    # Check they produce the right types.
    file_bin = await path.open("rb+")
    assert_type(await file_bin.read(), bytes)
    assert_type(await file_bin.write(b"test"), int)
    assert_type(await file_bin.seek(32), int)

    file_text = await path.open("r+t")
    assert_type(await file_text.read(), str)
    assert_type(await file_text.write("test"), int)
    # TODO: report mypy bug: equiv to https://github.com/microsoft/pyright/issues/6833
    assert_type(await file_text.readlines(), list[str])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\utils\datetime.py ===
# Copyright (c) 2010-2024 openpyxl

"""Manage Excel date weirdness."""

# Python stdlib imports
import datetime
from math import isnan
import re


# constants
MAC_EPOCH = datetime.datetime(1904, 1, 1)
WINDOWS_EPOCH = datetime.datetime(1899, 12, 30)
CALENDAR_WINDOWS_1900 = 2415018.5   # Julian date of WINDOWS_EPOCH
CALENDAR_MAC_1904 = 2416480.5       # Julian date of MAC_EPOCH
CALENDAR_WINDOWS_1900 = WINDOWS_EPOCH
CALENDAR_MAC_1904 = MAC_EPOCH
SECS_PER_DAY = 86400

ISO_FORMAT = '%Y-%m-%dT%H:%M:%SZ'
ISO_REGEX = re.compile(r'''
(?P<date>(?P<year>\d{4})-(?P<month>\d{2})-(?P<day>\d{2}))?T?
(?P<time>(?P<hour>\d{2}):(?P<minute>\d{2})(:(?P<second>\d{2})(?P<microsecond>\.\d{1,3})?)?)?Z?''',
                                       re.VERBOSE)
ISO_DURATION = re.compile(r'PT((?P<hours>\d+)H)?((?P<minutes>\d+)M)?((?P<seconds>\d+(\.\d{1,3})?)S)?')


def to_ISO8601(dt):
    """Convert from a datetime to a timestamp string."""
    if hasattr(dt, "microsecond") and dt.microsecond:
        return dt.isoformat(timespec="milliseconds")
    return dt.isoformat()


def from_ISO8601(formatted_string):
    """Convert from a timestamp string to a datetime object. According to
    18.17.4 in the specification the following ISO 8601 formats are
    supported.

    Dates B.1.1 and B.2.1
    Times B.1.2 and B.2.2
    Datetimes B.1.3 and B.2.3

    There is no concept of timedeltas in the specification, but Excel
    writes them (in strict OOXML mode), so these are also understood.
    """
    if not formatted_string:
        return None

    match = ISO_REGEX.match(formatted_string)
    if match and any(match.groups()):
        parts = match.groupdict(0)
        for key in ["year", "month", "day", "hour", "minute", "second"]:
            if parts[key]:
                parts[key] = int(parts[key])

        if parts["microsecond"]:
            parts["microsecond"] = int(float(parts['microsecond']) * 1_000_000)

        if not parts["date"]:
            dt = datetime.time(parts['hour'], parts['minute'], parts['second'], parts["microsecond"])
        elif not parts["time"]:
            dt = datetime.date(parts['year'], parts['month'], parts['day'])
        else:
            del parts["time"]
            del parts["date"]
            dt = datetime.datetime(**parts)
        return dt

    match = ISO_DURATION.match(formatted_string)
    if match and any(match.groups()):
        parts = match.groupdict(0)
        for key, val in parts.items():
            if val:
                parts[key] = float(val)
        return datetime.timedelta(**parts)

    raise ValueError("Invalid datetime value {}".format(formatted_string))


def to_excel(dt, epoch=WINDOWS_EPOCH):
    """Convert Python datetime to Excel serial"""
    if isinstance(dt, datetime.time):
        return time_to_days(dt)
    if isinstance(dt, datetime.timedelta):
        return timedelta_to_days(dt)
    if isnan(dt.year):  # Pandas supports Not a Date
        return

    if not hasattr(dt, "date"):
        dt = datetime.datetime.combine(dt, datetime.time())

    # rebase on epoch and adjust for < 1900-03-01
    days = (dt - epoch).days
    if 0 < days <= 60 and epoch == WINDOWS_EPOCH:
        days -= 1
    return days + time_to_days(dt)


def from_excel(value, epoch=WINDOWS_EPOCH, timedelta=False):
    """Convert Excel serial to Python datetime"""
    if value is None:
        return

    if timedelta:
        td = datetime.timedelta(days=value)
        if td.microseconds:
            # round to millisecond precision
            td = datetime.timedelta(seconds=td.total_seconds() // 1,
                                    microseconds=round(td.microseconds, -3))
        return td

    day, fraction = divmod(value, 1)
    diff = datetime.timedelta(milliseconds=round(fraction * SECS_PER_DAY * 1000))
    if 0 <= value < 1 and diff.days == 0:
        return days_to_time(diff)
    if 0 < value < 60 and epoch == WINDOWS_EPOCH:
        day += 1
    return epoch + datetime.timedelta(days=day) + diff


def time_to_days(value):
    """Convert a time value to fractions of day"""
    return (
        (value.hour * 3600)
        + (value.minute * 60)
        + value.second
        + value.microsecond / 10**6
        ) / SECS_PER_DAY


def timedelta_to_days(value):
    """Convert a timedelta value to fractions of a day"""
    return value.total_seconds() / SECS_PER_DAY


def days_to_time(value):
    mins, seconds = divmod(value.seconds, 60)
    hours, mins = divmod(mins, 60)
    return datetime.time(hours, mins, seconds, value.microseconds)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\api.py ===
from pandas._libs import (
    NaT,
    Period,
    Timedelta,
    Timestamp,
)
from pandas._libs.missing import NA

from pandas.core.dtypes.dtypes import (
    ArrowDtype,
    CategoricalDtype,
    DatetimeTZDtype,
    IntervalDtype,
    PeriodDtype,
)
from pandas.core.dtypes.missing import (
    isna,
    isnull,
    notna,
    notnull,
)

from pandas.core.algorithms import (
    factorize,
    unique,
    value_counts,
)
from pandas.core.arrays import Categorical
from pandas.core.arrays.boolean import BooleanDtype
from pandas.core.arrays.floating import (
    Float32Dtype,
    Float64Dtype,
)
from pandas.core.arrays.integer import (
    Int8Dtype,
    Int16Dtype,
    Int32Dtype,
    Int64Dtype,
    UInt8Dtype,
    UInt16Dtype,
    UInt32Dtype,
    UInt64Dtype,
)
from pandas.core.arrays.string_ import StringDtype
from pandas.core.construction import array
from pandas.core.flags import Flags
from pandas.core.groupby import (
    Grouper,
    NamedAgg,
)
from pandas.core.indexes.api import (
    CategoricalIndex,
    DatetimeIndex,
    Index,
    IntervalIndex,
    MultiIndex,
    PeriodIndex,
    RangeIndex,
    TimedeltaIndex,
)
from pandas.core.indexes.datetimes import (
    bdate_range,
    date_range,
)
from pandas.core.indexes.interval import (
    Interval,
    interval_range,
)
from pandas.core.indexes.period import period_range
from pandas.core.indexes.timedeltas import timedelta_range
from pandas.core.indexing import IndexSlice
from pandas.core.series import Series
from pandas.core.tools.datetimes import to_datetime
from pandas.core.tools.numeric import to_numeric
from pandas.core.tools.timedeltas import to_timedelta

from pandas.io.formats.format import set_eng_float_format
from pandas.tseries.offsets import DateOffset

# DataFrame needs to be imported after NamedAgg to avoid a circular import
from pandas.core.frame import DataFrame  # isort:skip

__all__ = [
    "array",
    "ArrowDtype",
    "bdate_range",
    "BooleanDtype",
    "Categorical",
    "CategoricalDtype",
    "CategoricalIndex",
    "DataFrame",
    "DateOffset",
    "date_range",
    "DatetimeIndex",
    "DatetimeTZDtype",
    "factorize",
    "Flags",
    "Float32Dtype",
    "Float64Dtype",
    "Grouper",
    "Index",
    "IndexSlice",
    "Int16Dtype",
    "Int32Dtype",
    "Int64Dtype",
    "Int8Dtype",
    "Interval",
    "IntervalDtype",
    "IntervalIndex",
    "interval_range",
    "isna",
    "isnull",
    "MultiIndex",
    "NA",
    "NamedAgg",
    "NaT",
    "notna",
    "notnull",
    "Period",
    "PeriodDtype",
    "PeriodIndex",
    "period_range",
    "RangeIndex",
    "Series",
    "set_eng_float_format",
    "StringDtype",
    "Timedelta",
    "TimedeltaIndex",
    "timedelta_range",
    "Timestamp",
    "to_datetime",
    "to_numeric",
    "to_timedelta",
    "UInt16Dtype",
    "UInt32Dtype",
    "UInt64Dtype",
    "UInt8Dtype",
    "unique",
    "value_counts",
]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\metadata\importlib\_envs.py ===
import importlib.metadata
import logging
import os
import pathlib
import sys
import zipfile
from typing import Iterator, List, Optional, Sequence, Set, Tuple

from pip._vendor.packaging.utils import (
    InvalidWheelFilename,
    NormalizedName,
    canonicalize_name,
    parse_wheel_filename,
)

from pip._internal.metadata.base import BaseDistribution, BaseEnvironment
from pip._internal.utils.filetypes import WHEEL_EXTENSION

from ._compat import BadMetadata, BasePath, get_dist_canonical_name, get_info_location
from ._dists import Distribution

logger = logging.getLogger(__name__)


def _looks_like_wheel(location: str) -> bool:
    if not location.endswith(WHEEL_EXTENSION):
        return False
    if not os.path.isfile(location):
        return False
    try:
        parse_wheel_filename(os.path.basename(location))
    except InvalidWheelFilename:
        return False
    return zipfile.is_zipfile(location)


class _DistributionFinder:
    """Finder to locate distributions.

    The main purpose of this class is to memoize found distributions' names, so
    only one distribution is returned for each package name. At lot of pip code
    assumes this (because it is setuptools's behavior), and not doing the same
    can potentially cause a distribution in lower precedence path to override a
    higher precedence one if the caller is not careful.

    Eventually we probably want to make it possible to see lower precedence
    installations as well. It's useful feature, after all.
    """

    FoundResult = Tuple[importlib.metadata.Distribution, Optional[BasePath]]

    def __init__(self) -> None:
        self._found_names: Set[NormalizedName] = set()

    def _find_impl(self, location: str) -> Iterator[FoundResult]:
        """Find distributions in a location."""
        # Skip looking inside a wheel. Since a package inside a wheel is not
        # always valid (due to .data directories etc.), its .dist-info entry
        # should not be considered an installed distribution.
        if _looks_like_wheel(location):
            return
        # To know exactly where we find a distribution, we have to feed in the
        # paths one by one, instead of dumping the list to importlib.metadata.
        for dist in importlib.metadata.distributions(path=[location]):
            info_location = get_info_location(dist)
            try:
                name = get_dist_canonical_name(dist)
            except BadMetadata as e:
                logger.warning("Skipping %s due to %s", info_location, e.reason)
                continue
            if name in self._found_names:
                continue
            self._found_names.add(name)
            yield dist, info_location

    def find(self, location: str) -> Iterator[BaseDistribution]:
        """Find distributions in a location.

        The path can be either a directory, or a ZIP archive.
        """
        for dist, info_location in self._find_impl(location):
            if info_location is None:
                installed_location: Optional[BasePath] = None
            else:
                installed_location = info_location.parent
            yield Distribution(dist, info_location, installed_location)

    def find_legacy_editables(self, location: str) -> Iterator[BaseDistribution]:
        """Read location in egg-link files and return distributions in there.

        The path should be a directory; otherwise this returns nothing. This
        follows how setuptools does this for compatibility. The first non-empty
        line in the egg-link is read as a path (resolved against the egg-link's
        containing directory if relative). Distributions found at that linked
        location are returned.
        """
        path = pathlib.Path(location)
        if not path.is_dir():
            return
        for child in path.iterdir():
            if child.suffix != ".egg-link":
                continue
            with child.open() as f:
                lines = (line.strip() for line in f)
                target_rel = next((line for line in lines if line), "")
            if not target_rel:
                continue
            target_location = str(path.joinpath(target_rel))
            for dist, info_location in self._find_impl(target_location):
                yield Distribution(dist, info_location, path)


class Environment(BaseEnvironment):
    def __init__(self, paths: Sequence[str]) -> None:
        self._paths = paths

    @classmethod
    def default(cls) -> BaseEnvironment:
        return cls(sys.path)

    @classmethod
    def from_paths(cls, paths: Optional[List[str]]) -> BaseEnvironment:
        if paths is None:
            return cls(sys.path)
        return cls(paths)

    def _iter_distributions(self) -> Iterator[BaseDistribution]:
        finder = _DistributionFinder()
        for location in self._paths:
            yield from finder.find(location)
            yield from finder.find_legacy_editables(location)

    def get_distribution(self, name: str) -> Optional[BaseDistribution]:
        canonical_name = canonicalize_name(name)
        matches = (
            distribution
            for distribution in self.iter_all_distributions()
            if distribution.canonical_name == canonical_name
        )
        return next(matches, None)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pytz\reference.py ===
'''
Reference tzinfo implementations from the Python docs.
Used for testing against as they are only correct for the years
1987 to 2006. Do not use these for real code.
'''

from datetime import tzinfo, timedelta, datetime
from pytz import HOUR, ZERO, UTC

__all__ = [
    'FixedOffset',
    'LocalTimezone',
    'USTimeZone',
    'Eastern',
    'Central',
    'Mountain',
    'Pacific',
    'UTC'
]


# A class building tzinfo objects for fixed-offset time zones.
# Note that FixedOffset(0, "UTC") is a different way to build a
# UTC tzinfo object.
class FixedOffset(tzinfo):
    """Fixed offset in minutes east from UTC."""

    def __init__(self, offset, name):
        self.__offset = timedelta(minutes=offset)
        self.__name = name

    def utcoffset(self, dt):
        return self.__offset

    def tzname(self, dt):
        return self.__name

    def dst(self, dt):
        return ZERO


import time as _time

STDOFFSET = timedelta(seconds=-_time.timezone)
if _time.daylight:
    DSTOFFSET = timedelta(seconds=-_time.altzone)
else:
    DSTOFFSET = STDOFFSET

DSTDIFF = DSTOFFSET - STDOFFSET


# A class capturing the platform's idea of local time.
class LocalTimezone(tzinfo):

    def utcoffset(self, dt):
        if self._isdst(dt):
            return DSTOFFSET
        else:
            return STDOFFSET

    def dst(self, dt):
        if self._isdst(dt):
            return DSTDIFF
        else:
            return ZERO

    def tzname(self, dt):
        return _time.tzname[self._isdst(dt)]

    def _isdst(self, dt):
        tt = (dt.year, dt.month, dt.day,
              dt.hour, dt.minute, dt.second,
              dt.weekday(), 0, -1)
        stamp = _time.mktime(tt)
        tt = _time.localtime(stamp)
        return tt.tm_isdst > 0

Local = LocalTimezone()


def first_sunday_on_or_after(dt):
    days_to_go = 6 - dt.weekday()
    if days_to_go:
        dt += timedelta(days_to_go)
    return dt


# In the US, DST starts at 2am (standard time) on the first Sunday in April.
DSTSTART = datetime(1, 4, 1, 2)
# and ends at 2am (DST time; 1am standard time) on the last Sunday of Oct.
# which is the first Sunday on or after Oct 25.
DSTEND = datetime(1, 10, 25, 1)


# A complete implementation of current DST rules for major US time zones.
class USTimeZone(tzinfo):

    def __init__(self, hours, reprname, stdname, dstname):
        self.stdoffset = timedelta(hours=hours)
        self.reprname = reprname
        self.stdname = stdname
        self.dstname = dstname

    def __repr__(self):
        return self.reprname

    def tzname(self, dt):
        if self.dst(dt):
            return self.dstname
        else:
            return self.stdname

    def utcoffset(self, dt):
        return self.stdoffset + self.dst(dt)

    def dst(self, dt):
        if dt is None or dt.tzinfo is None:
            # An exception may be sensible here, in one or both cases.
            # It depends on how you want to treat them.  The default
            # fromutc() implementation (called by the default astimezone()
            # implementation) passes a datetime with dt.tzinfo is self.
            return ZERO
        assert dt.tzinfo is self

        # Find first Sunday in April & the last in October.
        start = first_sunday_on_or_after(DSTSTART.replace(year=dt.year))
        end = first_sunday_on_or_after(DSTEND.replace(year=dt.year))

        # Can't compare naive to aware objects, so strip the timezone from
        # dt first.
        if start <= dt.replace(tzinfo=None) < end:
            return HOUR
        else:
            return ZERO

Eastern = USTimeZone(-5, "Eastern", "EST", "EDT")
Central = USTimeZone(-6, "Central", "CST", "CDT")
Mountain = USTimeZone(-7, "Mountain", "MST", "MDT")
Pacific = USTimeZone(-8, "Pacific", "PST", "PDT")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\tools\test_mypy_annotate.py ===
from __future__ import annotations

import io
import sys
from typing import TYPE_CHECKING

import pytest

from trio._tools.mypy_annotate import Result, export, main, process_line

if TYPE_CHECKING:
    from pathlib import Path


@pytest.mark.parametrize(
    ("src", "expected"),
    [
        ("", None),
        ("a regular line\n", None),
        (
            "package\\filename.py:42:8: note: Some info\n",
            Result(
                kind="notice",
                filename="package\\filename.py",
                start_line=42,
                start_col=8,
                end_line=None,
                end_col=None,
                message=" Some info",
            ),
        ),
        (
            "package/filename.py:42:1:46:3: error: Type error here [code]\n",
            Result(
                kind="error",
                filename="package/filename.py",
                start_line=42,
                start_col=1,
                end_line=46,
                end_col=3,
                message=" Type error here [code]",
            ),
        ),
        (
            "package/module.py:87: warn: Bad code\n",
            Result(
                kind="warning",
                filename="package/module.py",
                start_line=87,
                message=" Bad code",
            ),
        ),
    ],
    ids=["blank", "normal", "note-wcol", "error-wend", "warn-lineonly"],
)
def test_processing(src: str, expected: Result | None) -> None:
    result = process_line(src)
    assert result == expected


def test_export(capsys: pytest.CaptureFixture[str]) -> None:
    results = {
        Result(
            kind="notice",
            filename="package\\filename.py",
            start_line=42,
            start_col=8,
            end_line=None,
            end_col=None,
            message=" Some info",
        ): ["Windows", "Mac"],
        Result(
            kind="error",
            filename="package/filename.py",
            start_line=42,
            start_col=1,
            end_line=46,
            end_col=3,
            message=" Type error here [code]",
        ): ["Linux", "Mac"],
        Result(
            kind="warning",
            filename="package/module.py",
            start_line=87,
            message=" Bad code",
        ): ["Linux"],
    }
    export(results)
    std = capsys.readouterr()
    assert std.err == ""
    assert std.out == (
        "::notice file=package\\filename.py,line=42,col=8,"
        "title=Mypy-Windows+Mac::package\\filename.py:(42:8): Some info"
        "\n"
        "::error file=package/filename.py,line=42,col=1,endLine=46,endColumn=3,"
        "title=Mypy-Linux+Mac::package/filename.py:(42:1 - 46:3): Type error here [code]"
        "\n"
        "::warning file=package/module.py,line=87,"
        "title=Mypy-Linux::package/module.py:87: Bad code\n"
    )


def test_endtoend(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    import trio._tools.mypy_annotate as mypy_annotate

    inp_text = """\
Mypy begun
trio/core.py:15: error: Bad types here [misc]
trio/package/module.py:48:4:56:18: warn: Missing annotations  [no-untyped-def]
Found 3 errors in 29 files
"""
    result_file = tmp_path / "dump.dat"
    assert not result_file.exists()
    with monkeypatch.context():
        monkeypatch.setattr(sys, "stdin", io.StringIO(inp_text))

        mypy_annotate.main(
            ["--dumpfile", str(result_file), "--platform", "SomePlatform"],
        )

    std = capsys.readouterr()
    assert std.err == ""
    assert std.out == inp_text  # Echos the original.

    assert result_file.exists()

    main(["--dumpfile", str(result_file)])

    std = capsys.readouterr()
    assert std.err == ""
    assert std.out == (
        "::error file=trio/core.py,line=15,title=Mypy-SomePlatform::trio/core.py:15: Bad types here [misc]\n"
        "::warning file=trio/package/module.py,line=48,col=4,endLine=56,endColumn=18,"
        "title=Mypy-SomePlatform::trio/package/module.py:(48:4 - 56:18): Missing "
        "annotations  [no-untyped-def]\n"
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\models\wheel.py ===
"""Represents a wheel file and provides access to the various parts of the
name that have meaning.
"""

import re
from typing import Dict, Iterable, List, Optional

from pip._vendor.packaging.tags import Tag
from pip._vendor.packaging.utils import BuildTag, parse_wheel_filename
from pip._vendor.packaging.utils import (
    InvalidWheelFilename as _PackagingInvalidWheelFilename,
)

from pip._internal.exceptions import InvalidWheelFilename
from pip._internal.utils.deprecation import deprecated


class Wheel:
    """A wheel file"""

    legacy_wheel_file_re = re.compile(
        r"""^(?P<namever>(?P<name>[^\s-]+?)-(?P<ver>[^\s-]*?))
        ((-(?P<build>\d[^-]*?))?-(?P<pyver>[^\s-]+?)-(?P<abi>[^\s-]+?)-(?P<plat>[^\s-]+?)
        \.whl|\.dist-info)$""",
        re.VERBOSE,
    )

    def __init__(self, filename: str) -> None:
        self.filename = filename

        # To make mypy happy specify type hints that can come from either
        # parse_wheel_filename or the legacy_wheel_file_re match.
        self.name: str
        self._build_tag: Optional[BuildTag] = None

        try:
            wheel_info = parse_wheel_filename(filename)
            self.name, _version, self._build_tag, self.file_tags = wheel_info
            self.version = str(_version)
        except _PackagingInvalidWheelFilename as e:
            # Check if the wheel filename is in the legacy format
            legacy_wheel_info = self.legacy_wheel_file_re.match(filename)
            if not legacy_wheel_info:
                raise InvalidWheelFilename(e.args[0]) from None

            deprecated(
                reason=(
                    f"Wheel filename {filename!r} is not correctly normalised. "
                    "Future versions of pip will raise the following error:\n"
                    f"{e.args[0]}\n\n"
                ),
                replacement=(
                    "to rename the wheel to use a correctly normalised "
                    "name (this may require updating the version in "
                    "the project metadata)"
                ),
                gone_in="25.3",
                issue=12938,
            )

            self.name = legacy_wheel_info.group("name").replace("_", "-")
            self.version = legacy_wheel_info.group("ver").replace("_", "-")

            # Generate the file tags from the legacy wheel filename
            pyversions = legacy_wheel_info.group("pyver").split(".")
            abis = legacy_wheel_info.group("abi").split(".")
            plats = legacy_wheel_info.group("plat").split(".")
            self.file_tags = frozenset(
                Tag(interpreter=py, abi=abi, platform=plat)
                for py in pyversions
                for abi in abis
                for plat in plats
            )

    @property
    def build_tag(self) -> BuildTag:
        if self._build_tag is not None:
            return self._build_tag

        # Parse the build tag from the legacy wheel filename
        legacy_wheel_info = self.legacy_wheel_file_re.match(self.filename)
        assert legacy_wheel_info is not None, "guaranteed by filename validation"
        build_tag = legacy_wheel_info.group("build")
        match = re.match(r"^(\d+)(.*)$", build_tag)
        assert match is not None, "guaranteed by filename validation"
        build_tag_groups = match.groups()
        self._build_tag = (int(build_tag_groups[0]), build_tag_groups[1])

        return self._build_tag

    def get_formatted_file_tags(self) -> List[str]:
        """Return the wheel's tags as a sorted list of strings."""
        return sorted(str(tag) for tag in self.file_tags)

    def support_index_min(self, tags: List[Tag]) -> int:
        """Return the lowest index that one of the wheel's file_tag combinations
        achieves in the given list of supported tags.

        For example, if there are 8 supported tags and one of the file tags
        is first in the list, then return 0.

        :param tags: the PEP 425 tags to check the wheel against, in order
            with most preferred first.

        :raises ValueError: If none of the wheel's file tags match one of
            the supported tags.
        """
        try:
            return next(i for i, t in enumerate(tags) if t in self.file_tags)
        except StopIteration:
            raise ValueError()

    def find_most_preferred_tag(
        self, tags: List[Tag], tag_to_priority: Dict[Tag, int]
    ) -> int:
        """Return the priority of the most preferred tag that one of the wheel's file
        tag combinations achieves in the given list of supported tags using the given
        tag_to_priority mapping, where lower priorities are more-preferred.

        This is used in place of support_index_min in some cases in order to avoid
        an expensive linear scan of a large list of tags.

        :param tags: the PEP 425 tags to check the wheel against.
        :param tag_to_priority: a mapping from tag to priority of that tag, where
            lower is more preferred.

        :raises ValueError: If none of the wheel's file tags match one of
            the supported tags.
        """
        return min(
            tag_to_priority[tag] for tag in self.file_tags if tag in tag_to_priority
        )

    def supported(self, tags: Iterable[Tag]) -> bool:
        """Return whether the wheel is compatible with one of the given tags.

        :param tags: the PEP 425 tags to check the wheel against.
        """
        return not self.file_tags.isdisjoint(tags)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\resolution\resolvelib\base.py ===
from dataclasses import dataclass
from typing import FrozenSet, Iterable, Optional, Tuple

from pip._vendor.packaging.specifiers import SpecifierSet
from pip._vendor.packaging.utils import NormalizedName
from pip._vendor.packaging.version import Version

from pip._internal.models.link import Link, links_equivalent
from pip._internal.req.req_install import InstallRequirement
from pip._internal.utils.hashes import Hashes

CandidateLookup = Tuple[Optional["Candidate"], Optional[InstallRequirement]]


def format_name(project: NormalizedName, extras: FrozenSet[NormalizedName]) -> str:
    if not extras:
        return project
    extras_expr = ",".join(sorted(extras))
    return f"{project}[{extras_expr}]"


@dataclass(frozen=True)
class Constraint:
    specifier: SpecifierSet
    hashes: Hashes
    links: FrozenSet[Link]

    @classmethod
    def empty(cls) -> "Constraint":
        return Constraint(SpecifierSet(), Hashes(), frozenset())

    @classmethod
    def from_ireq(cls, ireq: InstallRequirement) -> "Constraint":
        links = frozenset([ireq.link]) if ireq.link else frozenset()
        return Constraint(ireq.specifier, ireq.hashes(trust_internet=False), links)

    def __bool__(self) -> bool:
        return bool(self.specifier) or bool(self.hashes) or bool(self.links)

    def __and__(self, other: InstallRequirement) -> "Constraint":
        if not isinstance(other, InstallRequirement):
            return NotImplemented
        specifier = self.specifier & other.specifier
        hashes = self.hashes & other.hashes(trust_internet=False)
        links = self.links
        if other.link:
            links = links.union([other.link])
        return Constraint(specifier, hashes, links)

    def is_satisfied_by(self, candidate: "Candidate") -> bool:
        # Reject if there are any mismatched URL constraints on this package.
        if self.links and not all(_match_link(link, candidate) for link in self.links):
            return False
        # We can safely always allow prereleases here since PackageFinder
        # already implements the prerelease logic, and would have filtered out
        # prerelease candidates if the user does not expect them.
        return self.specifier.contains(candidate.version, prereleases=True)


class Requirement:
    @property
    def project_name(self) -> NormalizedName:
        """The "project name" of a requirement.

        This is different from ``name`` if this requirement contains extras,
        in which case ``name`` would contain the ``[...]`` part, while this
        refers to the name of the project.
        """
        raise NotImplementedError("Subclass should override")

    @property
    def name(self) -> str:
        """The name identifying this requirement in the resolver.

        This is different from ``project_name`` if this requirement contains
        extras, where ``project_name`` would not contain the ``[...]`` part.
        """
        raise NotImplementedError("Subclass should override")

    def is_satisfied_by(self, candidate: "Candidate") -> bool:
        return False

    def get_candidate_lookup(self) -> CandidateLookup:
        raise NotImplementedError("Subclass should override")

    def format_for_error(self) -> str:
        raise NotImplementedError("Subclass should override")


def _match_link(link: Link, candidate: "Candidate") -> bool:
    if candidate.source_link:
        return links_equivalent(link, candidate.source_link)
    return False


class Candidate:
    @property
    def project_name(self) -> NormalizedName:
        """The "project name" of the candidate.

        This is different from ``name`` if this candidate contains extras,
        in which case ``name`` would contain the ``[...]`` part, while this
        refers to the name of the project.
        """
        raise NotImplementedError("Override in subclass")

    @property
    def name(self) -> str:
        """The name identifying this candidate in the resolver.

        This is different from ``project_name`` if this candidate contains
        extras, where ``project_name`` would not contain the ``[...]`` part.
        """
        raise NotImplementedError("Override in subclass")

    @property
    def version(self) -> Version:
        raise NotImplementedError("Override in subclass")

    @property
    def is_installed(self) -> bool:
        raise NotImplementedError("Override in subclass")

    @property
    def is_editable(self) -> bool:
        raise NotImplementedError("Override in subclass")

    @property
    def source_link(self) -> Optional[Link]:
        raise NotImplementedError("Override in subclass")

    def iter_dependencies(self, with_requires: bool) -> Iterable[Optional[Requirement]]:
        raise NotImplementedError("Override in subclass")

    def get_install_requirement(self) -> Optional[InstallRequirement]:
        raise NotImplementedError("Override in subclass")

    def format_for_error(self) -> str:
        raise NotImplementedError("Subclass should override")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\json.py ===
from pathlib import Path
from json import loads, dumps
from typing import Any, Callable, Optional, Union

from .text import Text
from .highlighter import JSONHighlighter, NullHighlighter


class JSON:
    """A renderable which pretty prints JSON.

    Args:
        json (str): JSON encoded data.
        indent (Union[None, int, str], optional): Number of characters to indent by. Defaults to 2.
        highlight (bool, optional): Enable highlighting. Defaults to True.
        skip_keys (bool, optional): Skip keys not of a basic type. Defaults to False.
        ensure_ascii (bool, optional): Escape all non-ascii characters. Defaults to False.
        check_circular (bool, optional): Check for circular references. Defaults to True.
        allow_nan (bool, optional): Allow NaN and Infinity values. Defaults to True.
        default (Callable, optional): A callable that converts values that can not be encoded
            in to something that can be JSON encoded. Defaults to None.
        sort_keys (bool, optional): Sort dictionary keys. Defaults to False.
    """

    def __init__(
        self,
        json: str,
        indent: Union[None, int, str] = 2,
        highlight: bool = True,
        skip_keys: bool = False,
        ensure_ascii: bool = False,
        check_circular: bool = True,
        allow_nan: bool = True,
        default: Optional[Callable[[Any], Any]] = None,
        sort_keys: bool = False,
    ) -> None:
        data = loads(json)
        json = dumps(
            data,
            indent=indent,
            skipkeys=skip_keys,
            ensure_ascii=ensure_ascii,
            check_circular=check_circular,
            allow_nan=allow_nan,
            default=default,
            sort_keys=sort_keys,
        )
        highlighter = JSONHighlighter() if highlight else NullHighlighter()
        self.text = highlighter(json)
        self.text.no_wrap = True
        self.text.overflow = None

    @classmethod
    def from_data(
        cls,
        data: Any,
        indent: Union[None, int, str] = 2,
        highlight: bool = True,
        skip_keys: bool = False,
        ensure_ascii: bool = False,
        check_circular: bool = True,
        allow_nan: bool = True,
        default: Optional[Callable[[Any], Any]] = None,
        sort_keys: bool = False,
    ) -> "JSON":
        """Encodes a JSON object from arbitrary data.

        Args:
            data (Any): An object that may be encoded in to JSON
            indent (Union[None, int, str], optional): Number of characters to indent by. Defaults to 2.
            highlight (bool, optional): Enable highlighting. Defaults to True.
            default (Callable, optional): Optional callable which will be called for objects that cannot be serialized. Defaults to None.
            skip_keys (bool, optional): Skip keys not of a basic type. Defaults to False.
            ensure_ascii (bool, optional): Escape all non-ascii characters. Defaults to False.
            check_circular (bool, optional): Check for circular references. Defaults to True.
            allow_nan (bool, optional): Allow NaN and Infinity values. Defaults to True.
            default (Callable, optional): A callable that converts values that can not be encoded
                in to something that can be JSON encoded. Defaults to None.
            sort_keys (bool, optional): Sort dictionary keys. Defaults to False.

        Returns:
            JSON: New JSON object from the given data.
        """
        json_instance: "JSON" = cls.__new__(cls)
        json = dumps(
            data,
            indent=indent,
            skipkeys=skip_keys,
            ensure_ascii=ensure_ascii,
            check_circular=check_circular,
            allow_nan=allow_nan,
            default=default,
            sort_keys=sort_keys,
        )
        highlighter = JSONHighlighter() if highlight else NullHighlighter()
        json_instance.text = highlighter(json)
        json_instance.text.no_wrap = True
        json_instance.text.overflow = None
        return json_instance

    def __rich__(self) -> Text:
        return self.text


if __name__ == "__main__":
    import argparse
    import sys

    parser = argparse.ArgumentParser(description="Pretty print json")
    parser.add_argument(
        "path",
        metavar="PATH",
        help="path to file, or - for stdin",
    )
    parser.add_argument(
        "-i",
        "--indent",
        metavar="SPACES",
        type=int,
        help="Number of spaces in an indent",
        default=2,
    )
    args = parser.parse_args()

    from pip._vendor.rich.console import Console

    console = Console()
    error_console = Console(stderr=True)

    try:
        if args.path == "-":
            json_data = sys.stdin.read()
        else:
            json_data = Path(args.path).read_text()
    except Exception as error:
        error_console.print(f"Unable to read {args.path!r}; {error}")
        sys.exit(-1)

    console.print(JSON(json_data, indent=args.indent), soft_wrap=True)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\selenium_manager.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.
import json
import logging
import os
import platform
import subprocess
import sys
import sysconfig
from pathlib import Path
from typing import List, Optional

from selenium.common import WebDriverException

logger = logging.getLogger(__name__)


class SeleniumManager:
    """Wrapper for getting information from the Selenium Manager binaries.

    This implementation is still in beta, and may change.
    """

    def binary_paths(self, args: List) -> dict:
        """Determines the locations of the requested assets.

        :Args:
         - args: the commands to send to the selenium manager binary.
        :Returns: dictionary of assets and their path
        """

        args = [str(self._get_binary())] + args
        if logger.getEffectiveLevel() == logging.DEBUG:
            args.append("--debug")
        args.append("--language-binding")
        args.append("python")
        args.append("--output")
        args.append("json")

        return self._run(args)

    @staticmethod
    def _get_binary() -> Path:
        """Determines the path of the correct Selenium Manager binary.

        :Returns: The Selenium Manager executable location

        :Raises: WebDriverException if the platform is unsupported
        """

        compiled_path = Path(__file__).parent.joinpath("selenium-manager")
        exe = sysconfig.get_config_var("EXE")
        if exe is not None:
            compiled_path = compiled_path.with_suffix(exe)

        path: Optional[Path] = None

        if (env_path := os.getenv("SE_MANAGER_PATH")) is not None:
            logger.debug("Selenium Manager set by env SE_MANAGER_PATH to: %s", env_path)
            path = Path(env_path)
        elif compiled_path.exists():
            path = compiled_path
        else:
            allowed = {
                ("darwin", "any"): "macos/selenium-manager",
                ("win32", "any"): "windows/selenium-manager.exe",
                ("cygwin", "any"): "windows/selenium-manager.exe",
                ("linux", "x86_64"): "linux/selenium-manager",
                ("freebsd", "x86_64"): "linux/selenium-manager",
                ("openbsd", "x86_64"): "linux/selenium-manager",
            }

            arch = platform.machine() if sys.platform in ("linux", "freebsd", "openbsd") else "any"
            if sys.platform in ["freebsd", "openbsd"]:
                logger.warning("Selenium Manager binary may not be compatible with %s; verify settings", sys.platform)

            location = allowed.get((sys.platform, arch))
            if location is None:
                raise WebDriverException(f"Unsupported platform/architecture combination: {sys.platform}/{arch}")

            path = Path(__file__).parent.joinpath(location)

        if path is None or not path.is_file():
            raise WebDriverException(f"Unable to obtain working Selenium Manager binary; {path}")

        logger.debug("Selenium Manager binary found at: %s", path)

        return path

    @staticmethod
    def _run(args: List[str]) -> dict:
        """Executes the Selenium Manager Binary.

        :Args:
         - args: the components of the command being executed.
        :Returns: The log string containing the driver location.
        """
        command = " ".join(args)
        logger.debug("Executing process: %s", command)
        try:
            if sys.platform == "win32":
                completed_proc = subprocess.run(args, capture_output=True, creationflags=subprocess.CREATE_NO_WINDOW)
            else:
                completed_proc = subprocess.run(args, capture_output=True)
            stdout = completed_proc.stdout.decode("utf-8").rstrip("\n")
            stderr = completed_proc.stderr.decode("utf-8").rstrip("\n")
            output = json.loads(stdout) if stdout != "" else {"logs": [], "result": {}}
        except Exception as err:
            raise WebDriverException(f"Unsuccessful command executed: {command}") from err

        SeleniumManager._process_logs(output["logs"])
        result = output["result"]
        if completed_proc.returncode:
            raise WebDriverException(
                f"Unsuccessful command executed: {command}; code: {completed_proc.returncode}\n{result}\n{stderr}"
            )
        return result

    @staticmethod
    def _process_logs(log_items: List[dict]):
        for item in log_items:
            if item["level"] == "WARN":
                logger.warning(item["message"])
            elif item["level"] in ["DEBUG", "INFO"]:
                logger.debug(item["message"])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\utils.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.
"""The Utils methods."""

import socket
from typing import Iterable, List, Optional, Union

from selenium.types import AnyKey
from selenium.webdriver.common.keys import Keys

_is_connectable_exceptions = (socket.error, ConnectionResetError)


def free_port() -> int:
    """Determines a free port using sockets."""
    free_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    free_socket.bind(("127.0.0.1", 0))
    free_socket.listen(5)
    port: int = free_socket.getsockname()[1]
    free_socket.close()
    return port


def find_connectable_ip(host: Union[str, bytes, bytearray, None], port: Optional[int] = None) -> Optional[str]:
    """Resolve a hostname to an IP, preferring IPv4 addresses.

    We prefer IPv4 so that we don't change behavior from previous IPv4-only
    implementations, and because some drivers (e.g., FirefoxDriver) do not
    support IPv6 connections.

    If the optional port number is provided, only IPs that listen on the given
    port are considered.

    :Args:
        - host - A hostname.
        - port - Optional port number.

    :Returns:
        A single IP address, as a string. If any IPv4 address is found, one is
        returned. Otherwise, if any IPv6 address is found, one is returned. If
        neither, then None is returned.
    """
    try:
        addrinfos = socket.getaddrinfo(host, None)
    except socket.gaierror:
        return None

    ip = None
    for family, _, _, _, sockaddr in addrinfos:
        connectable = True
        if port:
            connectable = is_connectable(port, sockaddr[0])

        if connectable and family == socket.AF_INET:
            return sockaddr[0]
        if connectable and not ip and family == socket.AF_INET6:
            ip = sockaddr[0]
    return ip


def join_host_port(host: str, port: int) -> str:
    """Joins a hostname and port together.

    This is a minimal implementation intended to cope with IPv6 literals. For
    example, _join_host_port('::1', 80) == '[::1]:80'.

    :Args:
        - host - A hostname.
        - port - An integer port.
    """
    if ":" in host and not host.startswith("["):
        return f"[{host}]:{port}"
    return f"{host}:{port}"


def is_connectable(port: int, host: Optional[str] = "localhost") -> bool:
    """Tries to connect to the server at port to see if it is running.

    :Args:
     - port - The port to connect.
    """
    socket_ = None
    try:
        socket_ = socket.create_connection((host, port), 1)
        result = True
    except _is_connectable_exceptions:
        result = False
    finally:
        if socket_:
            try:
                socket_.shutdown(socket.SHUT_RDWR)
            except Exception:
                pass
            socket_.close()
    return result


def is_url_connectable(port: Union[int, str]) -> bool:
    """Tries to connect to the HTTP server at /status path and specified port
    to see if it responds successfully.

    :Args:
     - port - The port to connect.
    """
    from urllib import request as url_request

    try:
        res = url_request.urlopen(f"http://127.0.0.1:{port}/status")
        return res.getcode() == 200
    except Exception:
        return False


def keys_to_typing(value: Iterable[AnyKey]) -> List[str]:
    """Processes the values that will be typed in the element."""
    characters: List[str] = []
    for val in value:
        if isinstance(val, Keys):
            # Todo: Does this even work?
            characters.append(val)
        elif isinstance(val, (int, float)):
            characters.extend(str(val))
        else:
            characters.extend(val)
    return characters

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v135\device_access.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: DeviceAccess (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

class RequestId(str):
    '''
    Device request id.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> RequestId:
        return cls(json)

    def __repr__(self):
        return 'RequestId({})'.format(super().__repr__())


class DeviceId(str):
    '''
    A device id.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> DeviceId:
        return cls(json)

    def __repr__(self):
        return 'DeviceId({})'.format(super().__repr__())


@dataclass
class PromptDevice:
    '''
    Device information displayed in a user prompt to select a device.
    '''
    id_: DeviceId

    #: Display name as it appears in a device request user prompt.
    name: str

    def to_json(self):
        json = dict()
        json['id'] = self.id_.to_json()
        json['name'] = self.name
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            id_=DeviceId.from_json(json['id']),
            name=str(json['name']),
        )


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enable events in this domain.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.enable',
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disable events in this domain.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.disable',
    }
    json = yield cmd_dict


def select_prompt(
        id_: RequestId,
        device_id: DeviceId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Select a device in response to a DeviceAccess.deviceRequestPrompted event.

    :param id_:
    :param device_id:
    '''
    params: T_JSON_DICT = dict()
    params['id'] = id_.to_json()
    params['deviceId'] = device_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.selectPrompt',
        'params': params,
    }
    json = yield cmd_dict


def cancel_prompt(
        id_: RequestId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Cancel a prompt in response to a DeviceAccess.deviceRequestPrompted event.

    :param id_:
    '''
    params: T_JSON_DICT = dict()
    params['id'] = id_.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.cancelPrompt',
        'params': params,
    }
    json = yield cmd_dict


@event_class('DeviceAccess.deviceRequestPrompted')
@dataclass
class DeviceRequestPrompted:
    '''
    A device request opened a user prompt to select a device. Respond with the
    selectPrompt or cancelPrompt command.
    '''
    id_: RequestId
    devices: typing.List[PromptDevice]

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DeviceRequestPrompted:
        return cls(
            id_=RequestId.from_json(json['id']),
            devices=[PromptDevice.from_json(i) for i in json['devices']]
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v136\device_access.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: DeviceAccess (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

class RequestId(str):
    '''
    Device request id.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> RequestId:
        return cls(json)

    def __repr__(self):
        return 'RequestId({})'.format(super().__repr__())


class DeviceId(str):
    '''
    A device id.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> DeviceId:
        return cls(json)

    def __repr__(self):
        return 'DeviceId({})'.format(super().__repr__())


@dataclass
class PromptDevice:
    '''
    Device information displayed in a user prompt to select a device.
    '''
    id_: DeviceId

    #: Display name as it appears in a device request user prompt.
    name: str

    def to_json(self):
        json = dict()
        json['id'] = self.id_.to_json()
        json['name'] = self.name
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            id_=DeviceId.from_json(json['id']),
            name=str(json['name']),
        )


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enable events in this domain.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.enable',
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disable events in this domain.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.disable',
    }
    json = yield cmd_dict


def select_prompt(
        id_: RequestId,
        device_id: DeviceId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Select a device in response to a DeviceAccess.deviceRequestPrompted event.

    :param id_:
    :param device_id:
    '''
    params: T_JSON_DICT = dict()
    params['id'] = id_.to_json()
    params['deviceId'] = device_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.selectPrompt',
        'params': params,
    }
    json = yield cmd_dict


def cancel_prompt(
        id_: RequestId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Cancel a prompt in response to a DeviceAccess.deviceRequestPrompted event.

    :param id_:
    '''
    params: T_JSON_DICT = dict()
    params['id'] = id_.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.cancelPrompt',
        'params': params,
    }
    json = yield cmd_dict


@event_class('DeviceAccess.deviceRequestPrompted')
@dataclass
class DeviceRequestPrompted:
    '''
    A device request opened a user prompt to select a device. Respond with the
    selectPrompt or cancelPrompt command.
    '''
    id_: RequestId
    devices: typing.List[PromptDevice]

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DeviceRequestPrompted:
        return cls(
            id_=RequestId.from_json(json['id']),
            devices=[PromptDevice.from_json(i) for i in json['devices']]
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v137\device_access.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: DeviceAccess (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

class RequestId(str):
    '''
    Device request id.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> RequestId:
        return cls(json)

    def __repr__(self):
        return 'RequestId({})'.format(super().__repr__())


class DeviceId(str):
    '''
    A device id.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> DeviceId:
        return cls(json)

    def __repr__(self):
        return 'DeviceId({})'.format(super().__repr__())


@dataclass
class PromptDevice:
    '''
    Device information displayed in a user prompt to select a device.
    '''
    id_: DeviceId

    #: Display name as it appears in a device request user prompt.
    name: str

    def to_json(self):
        json = dict()
        json['id'] = self.id_.to_json()
        json['name'] = self.name
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            id_=DeviceId.from_json(json['id']),
            name=str(json['name']),
        )


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enable events in this domain.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.enable',
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disable events in this domain.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.disable',
    }
    json = yield cmd_dict


def select_prompt(
        id_: RequestId,
        device_id: DeviceId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Select a device in response to a DeviceAccess.deviceRequestPrompted event.

    :param id_:
    :param device_id:
    '''
    params: T_JSON_DICT = dict()
    params['id'] = id_.to_json()
    params['deviceId'] = device_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.selectPrompt',
        'params': params,
    }
    json = yield cmd_dict


def cancel_prompt(
        id_: RequestId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Cancel a prompt in response to a DeviceAccess.deviceRequestPrompted event.

    :param id_:
    '''
    params: T_JSON_DICT = dict()
    params['id'] = id_.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DeviceAccess.cancelPrompt',
        'params': params,
    }
    json = yield cmd_dict


@event_class('DeviceAccess.deviceRequestPrompted')
@dataclass
class DeviceRequestPrompted:
    '''
    A device request opened a user prompt to select a device. Respond with the
    selectPrompt or cancelPrompt command.
    '''
    id_: RequestId
    devices: typing.List[PromptDevice]

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DeviceRequestPrompted:
        return cls(
            id_=RequestId.from_json(json['id']),
            devices=[PromptDevice.from_json(i) for i in json['devices']]
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\soupsieve\pretty.py ===
"""
Format a pretty string of a `SoupSieve` object for easy debugging.

This won't necessarily support all types and such, and definitely
not support custom outputs.

It is mainly geared towards our types as the `SelectorList`
object is a beast to look at without some indentation and newlines.
The format and various output types is fairly known (though it
hasn't been tested extensively to make sure we aren't missing corners).

Example:
-------
```
>>> import soupsieve as sv
>>> sv.compile('this > that.class[name=value]').selectors.pretty()
SelectorList(
    selectors=(
        Selector(
            tag=SelectorTag(
                name='that',
                prefix=None),
            ids=(),
            classes=(
                'class',
                ),
            attributes=(
                SelectorAttribute(
                    attribute='name',
                    prefix='',
                    pattern=re.compile(
                        '^value$'),
                    xml_type_pattern=None),
                ),
            nth=(),
            selectors=(),
            relation=SelectorList(
                selectors=(
                    Selector(
                        tag=SelectorTag(
                            name='this',
                            prefix=None),
                        ids=(),
                        classes=(),
                        attributes=(),
                        nth=(),
                        selectors=(),
                        relation=SelectorList(
                            selectors=(),
                            is_not=False,
                            is_html=False),
                        rel_type='>',
                        contains=(),
                        lang=(),
                        flags=0),
                    ),
                is_not=False,
                is_html=False),
            rel_type=None,
            contains=(),
            lang=(),
            flags=0),
        ),
    is_not=False,
    is_html=False)
```

"""
from __future__ import annotations
import re
from typing import Any

RE_CLASS = re.compile(r'(?i)[a-z_][_a-z\d\.]+\(')
RE_PARAM = re.compile(r'(?i)[_a-z][_a-z\d]+=')
RE_EMPTY = re.compile(r'\(\)|\[\]|\{\}')
RE_LSTRT = re.compile(r'\[')
RE_DSTRT = re.compile(r'\{')
RE_TSTRT = re.compile(r'\(')
RE_LEND = re.compile(r'\]')
RE_DEND = re.compile(r'\}')
RE_TEND = re.compile(r'\)')
RE_INT = re.compile(r'\d+')
RE_KWORD = re.compile(r'(?i)[_a-z][_a-z\d]+')
RE_DQSTR = re.compile(r'"(?:\\.|[^"\\])*"')
RE_SQSTR = re.compile(r"'(?:\\.|[^'\\])*'")
RE_SEP = re.compile(r'\s*(,)\s*')
RE_DSEP = re.compile(r'\s*(:)\s*')

TOKENS = {
    'class': RE_CLASS,
    'param': RE_PARAM,
    'empty': RE_EMPTY,
    'lstrt': RE_LSTRT,
    'dstrt': RE_DSTRT,
    'tstrt': RE_TSTRT,
    'lend': RE_LEND,
    'dend': RE_DEND,
    'tend': RE_TEND,
    'sqstr': RE_SQSTR,
    'sep': RE_SEP,
    'dsep': RE_DSEP,
    'int': RE_INT,
    'kword': RE_KWORD,
    'dqstr': RE_DQSTR
}


def pretty(obj: Any) -> str:  # pragma: no cover
    """Make the object output string pretty."""

    sel = str(obj)
    index = 0
    end = len(sel) - 1
    indent = 0
    output = []

    while index <= end:
        m = None
        for k, v in TOKENS.items():
            m = v.match(sel, index)

            if m:
                name = k
                index = m.end(0)
                if name in ('class', 'lstrt', 'dstrt', 'tstrt'):
                    indent += 4
                    output.append(f'{m.group(0)}\n{" " * indent}')
                elif name in ('param', 'int', 'kword', 'sqstr', 'dqstr', 'empty'):
                    output.append(m.group(0))
                elif name in ('lend', 'dend', 'tend'):
                    indent -= 4
                    output.append(m.group(0))
                elif name in ('sep',):
                    output.append(f'{m.group(1)}\n{" " * indent}')
                elif name in ('dsep',):
                    output.append(f'{m.group(1)} ')
                break

    return ''.join(output)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\mysql\pyodbc.py ===
# dialects/mysql/pyodbc.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors


r"""


.. dialect:: mysql+pyodbc
    :name: PyODBC
    :dbapi: pyodbc
    :connectstring: mysql+pyodbc://<username>:<password>@<dsnname>
    :url: https://pypi.org/project/pyodbc/

.. note::

    The PyODBC for MySQL dialect is **not tested as part of
    SQLAlchemy's continuous integration**.
    The recommended MySQL dialects are mysqlclient and PyMySQL.
    However, if you want to use the mysql+pyodbc dialect and require
    full support for ``utf8mb4`` characters (including supplementary
    characters like emoji) be sure to use a current release of
    MySQL Connector/ODBC and specify the "ANSI" (**not** "Unicode")
    version of the driver in your DSN or connection string.

Pass through exact pyodbc connection string::

    import urllib

    connection_string = (
        "DRIVER=MySQL ODBC 8.0 ANSI Driver;"
        "SERVER=localhost;"
        "PORT=3307;"
        "DATABASE=mydb;"
        "UID=root;"
        "PWD=(whatever);"
        "charset=utf8mb4;"
    )
    params = urllib.parse.quote_plus(connection_string)
    connection_uri = "mysql+pyodbc:///?odbc_connect=%s" % params

"""  # noqa

import re

from .base import MySQLDialect
from .base import MySQLExecutionContext
from .types import TIME
from ... import exc
from ... import util
from ...connectors.pyodbc import PyODBCConnector
from ...sql.sqltypes import Time


class _pyodbcTIME(TIME):
    def result_processor(self, dialect, coltype):
        def process(value):
            # pyodbc returns a datetime.time object; no need to convert
            return value

        return process


class MySQLExecutionContext_pyodbc(MySQLExecutionContext):
    def get_lastrowid(self):
        cursor = self.create_cursor()
        cursor.execute("SELECT LAST_INSERT_ID()")
        lastrowid = cursor.fetchone()[0]
        cursor.close()
        return lastrowid


class MySQLDialect_pyodbc(PyODBCConnector, MySQLDialect):
    supports_statement_cache = True
    colspecs = util.update_copy(MySQLDialect.colspecs, {Time: _pyodbcTIME})
    supports_unicode_statements = True
    execution_ctx_cls = MySQLExecutionContext_pyodbc

    pyodbc_driver_name = "MySQL"

    def _detect_charset(self, connection):
        """Sniff out the character set in use for connection results."""

        # Prefer 'character_set_results' for the current connection over the
        # value in the driver.  SET NAMES or individual variable SETs will
        # change the charset without updating the driver's view of the world.
        #
        # If it's decided that issuing that sort of SQL leaves you SOL, then
        # this can prefer the driver value.

        # set this to None as _fetch_setting attempts to use it (None is OK)
        self._connection_charset = None
        try:
            value = self._fetch_setting(connection, "character_set_client")
            if value:
                return value
        except exc.DBAPIError:
            pass

        util.warn(
            "Could not detect the connection character set.  "
            "Assuming latin1."
        )
        return "latin1"

    def _get_server_version_info(self, connection):
        return MySQLDialect._get_server_version_info(self, connection)

    def _extract_error_code(self, exception):
        m = re.compile(r"\((\d+)\)").search(str(exception.args))
        c = m.group(1)
        if c:
            return int(c)
        else:
            return None

    def on_connect(self):
        super_ = super().on_connect()

        def on_connect(conn):
            if super_ is not None:
                super_(conn)

            # declare Unicode encoding for pyodbc as per
            #   https://github.com/mkleehammer/pyodbc/wiki/Unicode
            pyodbc_SQL_CHAR = 1  # pyodbc.SQL_CHAR
            pyodbc_SQL_WCHAR = -8  # pyodbc.SQL_WCHAR
            conn.setdecoding(pyodbc_SQL_CHAR, encoding="utf-8")
            conn.setdecoding(pyodbc_SQL_WCHAR, encoding="utf-8")
            conn.setencoding(encoding="utf-8")

        return on_connect


dialect = MySQLDialect_pyodbc

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\testing\suite\test_update_delete.py ===
# testing/suite/test_update_delete.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors

from .. import fixtures
from ..assertions import eq_
from ..schema import Column
from ..schema import Table
from ... import Integer
from ... import String
from ... import testing


class SimpleUpdateDeleteTest(fixtures.TablesTest):
    run_deletes = "each"
    __requires__ = ("sane_rowcount",)
    __backend__ = True

    @classmethod
    def define_tables(cls, metadata):
        Table(
            "plain_pk",
            metadata,
            Column("id", Integer, primary_key=True),
            Column("data", String(50)),
        )

    @classmethod
    def insert_data(cls, connection):
        connection.execute(
            cls.tables.plain_pk.insert(),
            [
                {"id": 1, "data": "d1"},
                {"id": 2, "data": "d2"},
                {"id": 3, "data": "d3"},
            ],
        )

    def test_update(self, connection):
        t = self.tables.plain_pk
        r = connection.execute(
            t.update().where(t.c.id == 2), dict(data="d2_new")
        )
        assert not r.is_insert
        assert not r.returns_rows
        assert r.rowcount == 1

        eq_(
            connection.execute(t.select().order_by(t.c.id)).fetchall(),
            [(1, "d1"), (2, "d2_new"), (3, "d3")],
        )

    def test_delete(self, connection):
        t = self.tables.plain_pk
        r = connection.execute(t.delete().where(t.c.id == 2))
        assert not r.is_insert
        assert not r.returns_rows
        assert r.rowcount == 1
        eq_(
            connection.execute(t.select().order_by(t.c.id)).fetchall(),
            [(1, "d1"), (3, "d3")],
        )

    @testing.variation("criteria", ["rows", "norows", "emptyin"])
    @testing.requires.update_returning
    def test_update_returning(self, connection, criteria):
        t = self.tables.plain_pk

        stmt = t.update().returning(t.c.id, t.c.data)

        if criteria.norows:
            stmt = stmt.where(t.c.id == 10)
        elif criteria.rows:
            stmt = stmt.where(t.c.id == 2)
        elif criteria.emptyin:
            stmt = stmt.where(t.c.id.in_([]))
        else:
            criteria.fail()

        r = connection.execute(stmt, dict(data="d2_new"))
        assert not r.is_insert
        assert r.returns_rows
        eq_(r.keys(), ["id", "data"])

        if criteria.rows:
            eq_(r.all(), [(2, "d2_new")])
        else:
            eq_(r.all(), [])

        eq_(
            connection.execute(t.select().order_by(t.c.id)).fetchall(),
            (
                [(1, "d1"), (2, "d2_new"), (3, "d3")]
                if criteria.rows
                else [(1, "d1"), (2, "d2"), (3, "d3")]
            ),
        )

    @testing.variation("criteria", ["rows", "norows", "emptyin"])
    @testing.requires.delete_returning
    def test_delete_returning(self, connection, criteria):
        t = self.tables.plain_pk

        stmt = t.delete().returning(t.c.id, t.c.data)

        if criteria.norows:
            stmt = stmt.where(t.c.id == 10)
        elif criteria.rows:
            stmt = stmt.where(t.c.id == 2)
        elif criteria.emptyin:
            stmt = stmt.where(t.c.id.in_([]))
        else:
            criteria.fail()

        r = connection.execute(stmt)
        assert not r.is_insert
        assert r.returns_rows
        eq_(r.keys(), ["id", "data"])

        if criteria.rows:
            eq_(r.all(), [(2, "d2")])
        else:
            eq_(r.all(), [])

        eq_(
            connection.execute(t.select().order_by(t.c.id)).fetchall(),
            (
                [(1, "d1"), (3, "d3")]
                if criteria.rows
                else [(1, "d1"), (2, "d2"), (3, "d3")]
            ),
        )


__all__ = ("SimpleUpdateDeleteTest",)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\strategies\tree.py ===
from functools import partial
from sympy.strategies import chain, minimize
from sympy.strategies.core import identity
import sympy.strategies.branch as branch
from sympy.strategies.branch import yieldify


def treeapply(tree, join, leaf=identity):
    """ Apply functions onto recursive containers (tree).

    Explanation
    ===========

    join - a dictionary mapping container types to functions
      e.g. ``{list: minimize, tuple: chain}``

    Keys are containers/iterables.  Values are functions [a] -> a.

    Examples
    ========

    >>> from sympy.strategies.tree import treeapply
    >>> tree = [(3, 2), (4, 1)]
    >>> treeapply(tree, {list: max, tuple: min})
    2

    >>> add = lambda *args: sum(args)
    >>> def mul(*args):
    ...     total = 1
    ...     for arg in args:
    ...         total *= arg
    ...     return total
    >>> treeapply(tree, {list: mul, tuple: add})
    25
    """
    for typ in join:
        if isinstance(tree, typ):
            return join[typ](*map(partial(treeapply, join=join, leaf=leaf),
                                  tree))
    return leaf(tree)


def greedy(tree, objective=identity, **kwargs):
    """ Execute a strategic tree.  Select alternatives greedily

    Trees
    -----

    Nodes in a tree can be either

    function - a leaf
    list     - a selection among operations
    tuple    - a sequence of chained operations

    Textual examples
    ----------------

    Text: Run f, then run g, e.g. ``lambda x: g(f(x))``
    Code: ``(f, g)``

    Text: Run either f or g, whichever minimizes the objective
    Code: ``[f, g]``

    Textx: Run either f or g, whichever is better, then run h
    Code: ``([f, g], h)``

    Text: Either expand then simplify or try factor then foosimp. Finally print
    Code: ``([(expand, simplify), (factor, foosimp)], print)``

    Objective
    ---------

    "Better" is determined by the objective keyword.  This function makes
    choices to minimize the objective.  It defaults to the identity.

    Examples
    ========

    >>> from sympy.strategies.tree import greedy
    >>> inc    = lambda x: x + 1
    >>> dec    = lambda x: x - 1
    >>> double = lambda x: 2*x

    >>> tree = [inc, (dec, double)] # either inc or dec-then-double
    >>> fn = greedy(tree)
    >>> fn(4)  # lowest value comes from the inc
    5
    >>> fn(1)  # lowest value comes from dec then double
    0

    This function selects between options in a tuple.  The result is chosen
    that minimizes the objective function.

    >>> fn = greedy(tree, objective=lambda x: -x)  # maximize
    >>> fn(4)  # highest value comes from the dec then double
    6
    >>> fn(1)  # highest value comes from the inc
    2

    Greediness
    ----------

    This is a greedy algorithm.  In the example:

        ([a, b], c)  # do either a or b, then do c

    the choice between running ``a`` or ``b`` is made without foresight to c
    """
    optimize = partial(minimize, objective=objective)
    return treeapply(tree, {list: optimize, tuple: chain}, **kwargs)


def allresults(tree, leaf=yieldify):
    """ Execute a strategic tree.  Return all possibilities.

    Returns a lazy iterator of all possible results

    Exhaustiveness
    --------------

    This is an exhaustive algorithm.  In the example

        ([a, b], [c, d])

    All of the results from

        (a, c), (b, c), (a, d), (b, d)

    are returned.  This can lead to combinatorial blowup.

    See sympy.strategies.greedy for details on input
    """
    return treeapply(tree, {list: branch.multiplex, tuple: branch.chain},
                     leaf=leaf)


def brute(tree, objective=identity, **kwargs):
    return lambda expr: min(tuple(allresults(tree, **kwargs)(expr)),
                            key=objective)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\flatbuffers\table.py ===
# Copyright 2014 Google Inc. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from . import encode
from . import number_types as N


class Table(object):
    """Table wraps a byte slice and provides read access to its data.

    The variable `Pos` indicates the root of the FlatBuffers object therein."""

    __slots__ = ("Bytes", "Pos")

    def __init__(self, buf, pos):
        N.enforce_number(pos, N.UOffsetTFlags)

        self.Bytes = buf
        self.Pos = pos

    def Offset(self, vtableOffset):
        """Offset provides access into the Table's vtable.

        Deprecated fields are ignored by checking the vtable's length."""

        vtable = self.Pos - self.Get(N.SOffsetTFlags, self.Pos)
        vtableEnd = self.Get(N.VOffsetTFlags, vtable)
        if vtableOffset < vtableEnd:
            return self.Get(N.VOffsetTFlags, vtable + vtableOffset)
        return 0

    def Indirect(self, off):
        """Indirect retrieves the relative offset stored at `offset`."""
        N.enforce_number(off, N.UOffsetTFlags)
        return off + encode.Get(N.UOffsetTFlags.packer_type, self.Bytes, off)

    def String(self, off):
        """String gets a string from data stored inside the flatbuffer."""
        N.enforce_number(off, N.UOffsetTFlags)
        off += encode.Get(N.UOffsetTFlags.packer_type, self.Bytes, off)
        start = off + N.UOffsetTFlags.bytewidth
        length = encode.Get(N.UOffsetTFlags.packer_type, self.Bytes, off)
        return bytes(self.Bytes[start:start+length])

    def VectorLen(self, off):
        """VectorLen retrieves the length of the vector whose offset is stored
           at "off" in this object."""
        N.enforce_number(off, N.UOffsetTFlags)

        off += self.Pos
        off += encode.Get(N.UOffsetTFlags.packer_type, self.Bytes, off)
        ret = encode.Get(N.UOffsetTFlags.packer_type, self.Bytes, off)
        return ret

    def Vector(self, off):
        """Vector retrieves the start of data of the vector whose offset is
           stored at "off" in this object."""
        N.enforce_number(off, N.UOffsetTFlags)

        off += self.Pos
        x = off + self.Get(N.UOffsetTFlags, off)
        # data starts after metadata containing the vector length
        x += N.UOffsetTFlags.bytewidth
        return x

    def Union(self, t2, off):
        """Union initializes any Table-derived type to point to the union at
           the given offset."""
        assert type(t2) is Table
        N.enforce_number(off, N.UOffsetTFlags)

        off += self.Pos
        t2.Pos = off + self.Get(N.UOffsetTFlags, off)
        t2.Bytes = self.Bytes

    def Get(self, flags, off):
        """
        Get retrieves a value of the type specified by `flags`  at the
        given offset.
        """
        N.enforce_number(off, N.UOffsetTFlags)
        return flags.py_type(encode.Get(flags.packer_type, self.Bytes, off))

    def GetSlot(self, slot, d, validator_flags):
        N.enforce_number(slot, N.VOffsetTFlags)
        if validator_flags is not None:
            N.enforce_number(d, validator_flags)
        off = self.Offset(slot)
        if off == 0:
            return d
        return self.Get(validator_flags, self.Pos + off)

    def GetVectorAsNumpy(self, flags, off):
        """
        GetVectorAsNumpy returns the vector that starts at `Vector(off)`
        as a numpy array with the type specified by `flags`. The array is
        a `view` into Bytes, so modifying the returned array will
        modify Bytes in place.
        """
        offset = self.Vector(off)
        length = self.VectorLen(off) # TODO: length accounts for bytewidth, right?
        numpy_dtype = N.to_numpy_type(flags)
        return encode.GetVectorAsNumpy(numpy_dtype, self.Bytes, length, offset)

    def GetArrayAsNumpy(self, flags, off, length):
        """
        GetArrayAsNumpy returns the array with fixed width that starts at `Vector(offset)`
        with length `length` as a numpy array with the type specified by `flags`. The
        array is a `view` into Bytes so modifying the returned will modify Bytes in place.
        """
        numpy_dtype = N.to_numpy_type(flags)
        return encode.GetVectorAsNumpy(numpy_dtype, self.Bytes, length, off)

    def GetVOffsetTSlot(self, slot, d):
        """
        GetVOffsetTSlot retrieves the VOffsetT that the given vtable location
        points to. If the vtable value is zero, the default value `d`
        will be returned.
        """

        N.enforce_number(slot, N.VOffsetTFlags)
        N.enforce_number(d, N.VOffsetTFlags)

        off = self.Offset(slot)
        if off == 0:
            return d
        return off

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\doc\ufuncs.py ===
"""
===================
Universal Functions
===================

Ufuncs are, generally speaking, mathematical functions or operations that are
applied element-by-element to the contents of an array. That is, the result
in each output array element only depends on the value in the corresponding
input array (or arrays) and on no other array elements. NumPy comes with a
large suite of ufuncs, and scipy extends that suite substantially. The simplest
example is the addition operator: ::

 >>> np.array([0,2,3,4]) + np.array([1,1,-1,2])
 array([1, 3, 2, 6])

The ufunc module lists all the available ufuncs in numpy. Documentation on
the specific ufuncs may be found in those modules. This documentation is
intended to address the more general aspects of ufuncs common to most of
them. All of the ufuncs that make use of Python operators (e.g., +, -, etc.)
have equivalent functions defined (e.g. add() for +)

Type coercion
=============

What happens when a binary operator (e.g., +,-,\\*,/, etc) deals with arrays of
two different types? What is the type of the result? Typically, the result is
the higher of the two types. For example: ::

 float32 + float64 -> float64
 int8 + int32 -> int32
 int16 + float32 -> float32
 float32 + complex64 -> complex64

There are some less obvious cases generally involving mixes of types
(e.g. uints, ints and floats) where equal bit sizes for each are not
capable of saving all the information in a different type of equivalent
bit size. Some examples are int32 vs float32 or uint32 vs int32.
Generally, the result is the higher type of larger size than both
(if available). So: ::

 int32 + float32 -> float64
 uint32 + int32 -> int64

Finally, the type coercion behavior when expressions involve Python
scalars is different than that seen for arrays. Since Python has a
limited number of types, combining a Python int with a dtype=np.int8
array does not coerce to the higher type but instead, the type of the
array prevails. So the rules for Python scalars combined with arrays is
that the result will be that of the array equivalent the Python scalar
if the Python scalar is of a higher 'kind' than the array (e.g., float
vs. int), otherwise the resultant type will be that of the array.
For example: ::

  Python int + int8 -> int8
  Python float + int8 -> float64

ufunc methods
=============

Binary ufuncs support 4 methods.

**.reduce(arr)** applies the binary operator to elements of the array in
  sequence. For example: ::

 >>> np.add.reduce(np.arange(10))  # adds all elements of array
 45

For multidimensional arrays, the first dimension is reduced by default: ::

 >>> np.add.reduce(np.arange(10).reshape(2,5))
     array([ 5,  7,  9, 11, 13])

The axis keyword can be used to specify different axes to reduce: ::

 >>> np.add.reduce(np.arange(10).reshape(2,5),axis=1)
 array([10, 35])

**.accumulate(arr)** applies the binary operator and generates an
equivalently shaped array that includes the accumulated amount for each
element of the array. A couple examples: ::

 >>> np.add.accumulate(np.arange(10))
 array([ 0,  1,  3,  6, 10, 15, 21, 28, 36, 45])
 >>> np.multiply.accumulate(np.arange(1,9))
 array([    1,     2,     6,    24,   120,   720,  5040, 40320])

The behavior for multidimensional arrays is the same as for .reduce(),
as is the use of the axis keyword).

**.reduceat(arr,indices)** allows one to apply reduce to selected parts
  of an array. It is a difficult method to understand. See the documentation
  at:

**.outer(arr1,arr2)** generates an outer operation on the two arrays arr1 and
  arr2. It will work on multidimensional arrays (the shape of the result is
  the concatenation of the two input shapes.: ::

 >>> np.multiply.outer(np.arange(3),np.arange(4))
 array([[0, 0, 0, 0],
        [0, 1, 2, 3],
        [0, 2, 4, 6]])

Output arguments
================

All ufuncs accept an optional output array. The array must be of the expected
output shape. Beware that if the type of the output array is of a different
(and lower) type than the output result, the results may be silently truncated
or otherwise corrupted in the downcast to the lower type. This usage is useful
when one wants to avoid creating large temporary arrays and instead allows one
to reuse the same array memory repeatedly (at the expense of not being able to
use more convenient operator notation in expressions). Note that when the
output argument is used, the ufunc still returns a reference to the result.

 >>> x = np.arange(2)
 >>> np.add(np.arange(2, dtype=float), np.arange(2, dtype=float), x,
 ...        casting='unsafe')
 array([0, 2])
 >>> x
 array([0, 2])

and & or as ufuncs
==================

Invariably people try to use the python 'and' and 'or' as logical operators
(and quite understandably). But these operators do not behave as normal
operators since Python treats these quite differently. They cannot be
overloaded with array equivalents. Thus using 'and' or 'or' with an array
results in an error. There are two alternatives:

 1) use the ufunc functions logical_and() and logical_or().
 2) use the bitwise operators & and \\|. The drawback of these is that if
    the arguments to these operators are not boolean arrays, the result is
    likely incorrect. On the other hand, most usages of logical_and and
    logical_or are with boolean arrays. As long as one is careful, this is
    a convenient way to apply these operators.

"""

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\operations\build\build_tracker.py ===
import contextlib
import hashlib
import logging
import os
from types import TracebackType
from typing import Dict, Generator, Optional, Type, Union

from pip._internal.req.req_install import InstallRequirement
from pip._internal.utils.temp_dir import TempDirectory

logger = logging.getLogger(__name__)


@contextlib.contextmanager
def update_env_context_manager(**changes: str) -> Generator[None, None, None]:
    target = os.environ

    # Save values from the target and change them.
    non_existent_marker = object()
    saved_values: Dict[str, Union[object, str]] = {}
    for name, new_value in changes.items():
        try:
            saved_values[name] = target[name]
        except KeyError:
            saved_values[name] = non_existent_marker
        target[name] = new_value

    try:
        yield
    finally:
        # Restore original values in the target.
        for name, original_value in saved_values.items():
            if original_value is non_existent_marker:
                del target[name]
            else:
                assert isinstance(original_value, str)  # for mypy
                target[name] = original_value


@contextlib.contextmanager
def get_build_tracker() -> Generator["BuildTracker", None, None]:
    root = os.environ.get("PIP_BUILD_TRACKER")
    with contextlib.ExitStack() as ctx:
        if root is None:
            root = ctx.enter_context(TempDirectory(kind="build-tracker")).path
            ctx.enter_context(update_env_context_manager(PIP_BUILD_TRACKER=root))
            logger.debug("Initialized build tracking at %s", root)

        with BuildTracker(root) as tracker:
            yield tracker


class TrackerId(str):
    """Uniquely identifying string provided to the build tracker."""


class BuildTracker:
    """Ensure that an sdist cannot request itself as a setup requirement.

    When an sdist is prepared, it identifies its setup requirements in the
    context of ``BuildTracker.track()``. If a requirement shows up recursively, this
    raises an exception.

    This stops fork bombs embedded in malicious packages."""

    def __init__(self, root: str) -> None:
        self._root = root
        self._entries: Dict[TrackerId, InstallRequirement] = {}
        logger.debug("Created build tracker: %s", self._root)

    def __enter__(self) -> "BuildTracker":
        logger.debug("Entered build tracker: %s", self._root)
        return self

    def __exit__(
        self,
        exc_type: Optional[Type[BaseException]],
        exc_val: Optional[BaseException],
        exc_tb: Optional[TracebackType],
    ) -> None:
        self.cleanup()

    def _entry_path(self, key: TrackerId) -> str:
        hashed = hashlib.sha224(key.encode()).hexdigest()
        return os.path.join(self._root, hashed)

    def add(self, req: InstallRequirement, key: TrackerId) -> None:
        """Add an InstallRequirement to build tracking."""

        # Get the file to write information about this requirement.
        entry_path = self._entry_path(key)

        # Try reading from the file. If it exists and can be read from, a build
        # is already in progress, so a LookupError is raised.
        try:
            with open(entry_path) as fp:
                contents = fp.read()
        except FileNotFoundError:
            pass
        else:
            message = f"{req.link} is already being built: {contents}"
            raise LookupError(message)

        # If we're here, req should really not be building already.
        assert key not in self._entries

        # Start tracking this requirement.
        with open(entry_path, "w", encoding="utf-8") as fp:
            fp.write(str(req))
        self._entries[key] = req

        logger.debug("Added %s to build tracker %r", req, self._root)

    def remove(self, req: InstallRequirement, key: TrackerId) -> None:
        """Remove an InstallRequirement from build tracking."""

        # Delete the created file and the corresponding entry.
        os.unlink(self._entry_path(key))
        del self._entries[key]

        logger.debug("Removed %s from build tracker %r", req, self._root)

    def cleanup(self) -> None:
        for key, req in list(self._entries.items()):
            self.remove(req, key)

        logger.debug("Removed build tracker: %r", self._root)

    @contextlib.contextmanager
    def track(self, req: InstallRequirement, key: str) -> Generator[None, None, None]:
        """Ensure that `key` cannot install itself as a setup requirement.

        :raises LookupError: If `key` was already provided in a parent invocation of
                             the context introduced by this method."""
        tracker_id = TrackerId(key)
        self.add(req, tracker_id)
        yield
        self.remove(req, tracker_id)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\spinner.py ===
from typing import cast, List, Optional, TYPE_CHECKING, Union

from ._spinners import SPINNERS
from .measure import Measurement
from .table import Table
from .text import Text

if TYPE_CHECKING:
    from .console import Console, ConsoleOptions, RenderResult, RenderableType
    from .style import StyleType


class Spinner:
    """A spinner animation.

    Args:
        name (str): Name of spinner (run python -m rich.spinner).
        text (RenderableType, optional): A renderable to display at the right of the spinner (str or Text typically). Defaults to "".
        style (StyleType, optional): Style for spinner animation. Defaults to None.
        speed (float, optional): Speed factor for animation. Defaults to 1.0.

    Raises:
        KeyError: If name isn't one of the supported spinner animations.
    """

    def __init__(
        self,
        name: str,
        text: "RenderableType" = "",
        *,
        style: Optional["StyleType"] = None,
        speed: float = 1.0,
    ) -> None:
        try:
            spinner = SPINNERS[name]
        except KeyError:
            raise KeyError(f"no spinner called {name!r}")
        self.text: "Union[RenderableType, Text]" = (
            Text.from_markup(text) if isinstance(text, str) else text
        )
        self.name = name
        self.frames = cast(List[str], spinner["frames"])[:]
        self.interval = cast(float, spinner["interval"])
        self.start_time: Optional[float] = None
        self.style = style
        self.speed = speed
        self.frame_no_offset: float = 0.0
        self._update_speed = 0.0

    def __rich_console__(
        self, console: "Console", options: "ConsoleOptions"
    ) -> "RenderResult":
        yield self.render(console.get_time())

    def __rich_measure__(
        self, console: "Console", options: "ConsoleOptions"
    ) -> Measurement:
        text = self.render(0)
        return Measurement.get(console, options, text)

    def render(self, time: float) -> "RenderableType":
        """Render the spinner for a given time.

        Args:
            time (float): Time in seconds.

        Returns:
            RenderableType: A renderable containing animation frame.
        """
        if self.start_time is None:
            self.start_time = time

        frame_no = ((time - self.start_time) * self.speed) / (
            self.interval / 1000.0
        ) + self.frame_no_offset
        frame = Text(
            self.frames[int(frame_no) % len(self.frames)], style=self.style or ""
        )

        if self._update_speed:
            self.frame_no_offset = frame_no
            self.start_time = time
            self.speed = self._update_speed
            self._update_speed = 0.0

        if not self.text:
            return frame
        elif isinstance(self.text, (str, Text)):
            return Text.assemble(frame, " ", self.text)
        else:
            table = Table.grid(padding=1)
            table.add_row(frame, self.text)
            return table

    def update(
        self,
        *,
        text: "RenderableType" = "",
        style: Optional["StyleType"] = None,
        speed: Optional[float] = None,
    ) -> None:
        """Updates attributes of a spinner after it has been started.

        Args:
            text (RenderableType, optional): A renderable to display at the right of the spinner (str or Text typically). Defaults to "".
            style (StyleType, optional): Style for spinner animation. Defaults to None.
            speed (float, optional): Speed factor for animation. Defaults to None.
        """
        if text:
            self.text = Text.from_markup(text) if isinstance(text, str) else text
        if style:
            self.style = style
        if speed:
            self._update_speed = speed


if __name__ == "__main__":  # pragma: no cover
    from time import sleep

    from .columns import Columns
    from .panel import Panel
    from .live import Live

    all_spinners = Columns(
        [
            Spinner(spinner_name, text=Text(repr(spinner_name), style="green"))
            for spinner_name in sorted(SPINNERS.keys())
        ],
        column_first=True,
        expand=True,
    )

    with Live(
        Panel(all_spinners, title="Spinners", border_style="blue"),
        refresh_per_second=20,
    ) as live:
        while True:
            sleep(0.1)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pyreadline3\keysyms\keysyms.py ===
# -*- coding: utf-8 -*-
# *****************************************************************************
#       Copyright (C) 2003-2006 Gary Bishop.
#       Copyright (C) 2006-2020 Jorgen Stenarson. <jorgen.stenarson@bostream.nu>
#       Copyright (C) 2020 Bassem Girgis. <brgirgis@gmail.com>
#
#  Distributed under the terms of the BSD License.  The full license is in
#  the file COPYING, distributed as part of this software.
# *****************************************************************************


from ctypes import windll

from . import winconstants as c32
from .common import KeyPress

# table for translating virtual keys to X windows key symbols

code2sym_map = {
    c32.VK_CANCEL: "cancel",
    c32.VK_BACK: "backspace",
    c32.VK_TAB: "tab",
    c32.VK_CLEAR: "clear",
    c32.VK_RETURN: "return",
    c32.VK_SHIFT: "shift_l",
    c32.VK_CONTROL: "control_l",
    c32.VK_MENU: "alt_l",
    c32.VK_PAUSE: "pause",
    c32.VK_CAPITAL: "caps_lock",
    c32.VK_ESCAPE: "escape",
    c32.VK_SPACE: "space",
    c32.VK_PRIOR: "prior",
    c32.VK_NEXT: "next",
    c32.VK_END: "end",
    c32.VK_HOME: "home",
    c32.VK_LEFT: "left",
    c32.VK_UP: "up",
    c32.VK_RIGHT: "right",
    c32.VK_DOWN: "down",
    c32.VK_SELECT: "select",
    c32.VK_PRINT: "print",
    c32.VK_EXECUTE: "execute",
    c32.VK_SNAPSHOT: "snapshot",
    c32.VK_INSERT: "insert",
    c32.VK_DELETE: "delete",
    c32.VK_HELP: "help",
    c32.VK_F1: "f1",
    c32.VK_F2: "f2",
    c32.VK_F3: "f3",
    c32.VK_F4: "f4",
    c32.VK_F5: "f5",
    c32.VK_F6: "f6",
    c32.VK_F7: "f7",
    c32.VK_F8: "f8",
    c32.VK_F9: "f9",
    c32.VK_F10: "f10",
    c32.VK_F11: "f11",
    c32.VK_F12: "f12",
    c32.VK_F13: "f13",
    c32.VK_F14: "f14",
    c32.VK_F15: "f15",
    c32.VK_F16: "f16",
    c32.VK_F17: "f17",
    c32.VK_F18: "f18",
    c32.VK_F19: "f19",
    c32.VK_F20: "f20",
    c32.VK_F21: "f21",
    c32.VK_F22: "f22",
    c32.VK_F23: "f23",
    c32.VK_F24: "f24",
    c32.VK_NUMLOCK: "num_lock,",
    c32.VK_SCROLL: "scroll_lock",
    c32.VK_APPS: "vk_apps",
    c32.VK_PROCESSKEY: "vk_processkey",
    c32.VK_ATTN: "vk_attn",
    c32.VK_CRSEL: "vk_crsel",
    c32.VK_EXSEL: "vk_exsel",
    c32.VK_EREOF: "vk_ereof",
    c32.VK_PLAY: "vk_play",
    c32.VK_ZOOM: "vk_zoom",
    c32.VK_NONAME: "vk_noname",
    c32.VK_PA1: "vk_pa1",
    c32.VK_OEM_CLEAR: "vk_oem_clear",
    c32.VK_NUMPAD0: "numpad0",
    c32.VK_NUMPAD1: "numpad1",
    c32.VK_NUMPAD2: "numpad2",
    c32.VK_NUMPAD3: "numpad3",
    c32.VK_NUMPAD4: "numpad4",
    c32.VK_NUMPAD5: "numpad5",
    c32.VK_NUMPAD6: "numpad6",
    c32.VK_NUMPAD7: "numpad7",
    c32.VK_NUMPAD8: "numpad8",
    c32.VK_NUMPAD9: "numpad9",
    c32.VK_DIVIDE: "divide",
    c32.VK_MULTIPLY: "multiply",
    c32.VK_ADD: "add",
    c32.VK_SUBTRACT: "subtract",
    c32.VK_DECIMAL: "vk_decimal",
}

VkKeyScan = windll.user32.VkKeyScanA


def char_to_keyinfo(char, control=False, meta=False, shift=False):
    k = KeyPress()
    vk = VkKeyScan(ord(char))
    if vk & 0xFFFF == 0xFFFF:
        print('VkKeyScan("%s") = %x' % (char, vk))
        raise ValueError("bad key")
    if vk & 0x100:
        k.shift = True
    if vk & 0x200:
        k.control = True
    if vk & 0x400:
        k.meta = True
    k.char = chr(vk & 0xFF)
    return k


def make_KeyPress(char, state, keycode):
    control = (state & (4 + 8)) != 0
    meta = (state & (1 + 2)) != 0
    shift = (state & 0x10) != 0
    if control and not meta:  # Matches ctrl- chords should pass keycode as char
        char = chr(keycode)
    elif control and meta:  # Matches alt gr and should just pass on char
        control = False
        meta = False
    try:
        keyname = code2sym_map[keycode]
    except KeyError:
        keyname = ""
    out = KeyPress(char, shift, control, meta, keyname)
    return out


if __name__ == "__main__":
    import startup

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\quantum\innerproduct.py ===
"""Symbolic inner product."""

from sympy.core.expr import Expr
from sympy.core.kind import NumberKind
from sympy.functions.elementary.complexes import conjugate
from sympy.printing.pretty.stringpict import prettyForm
from sympy.physics.quantum.dagger import Dagger


__all__ = [
    'InnerProduct'
]


# InnerProduct is not an QExpr because it is really just a regular commutative
# number. We have gone back and forth about this, but we gain a lot by having
# it subclass Expr. The main challenges were getting Dagger to work
# (we use _eval_conjugate) and represent (we can use atoms and subs). Having
# it be an Expr, mean that there are no commutative QExpr subclasses,
# which simplifies the design of everything.

class InnerProduct(Expr):
    """An unevaluated inner product between a Bra and a Ket [1].

    Parameters
    ==========

    bra : BraBase or subclass
        The bra on the left side of the inner product.
    ket : KetBase or subclass
        The ket on the right side of the inner product.

    Examples
    ========

    Create an InnerProduct and check its properties:

        >>> from sympy.physics.quantum import Bra, Ket
        >>> b = Bra('b')
        >>> k = Ket('k')
        >>> ip = b*k
        >>> ip
        <b|k>
        >>> ip.bra
        <b|
        >>> ip.ket
        |k>

    In quantum expressions, inner products will be automatically
    identified and created::

        >>> b*k
        <b|k>

    In more complex expressions, where there is ambiguity in whether inner or
    outer products should be created, inner products have high priority::

        >>> k*b*k*b
        <b|k>*|k><b|

    Notice how the inner product <b|k> moved to the left of the expression
    because inner products are commutative complex numbers.

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Inner_product
    """

    kind = NumberKind

    is_complex = True

    def __new__(cls, bra, ket):
        # Keep the import of BraBase and KetBase here to avoid problems
        # with circular imports.
        from sympy.physics.quantum.state import KetBase, BraBase
        if not isinstance(ket, KetBase):
            raise TypeError('KetBase subclass expected, got: %r' % ket)
        if not isinstance(bra, BraBase):
            raise TypeError('BraBase subclass expected, got: %r' % ket)
        obj = Expr.__new__(cls, bra, ket)
        return obj

    @property
    def bra(self):
        return self.args[0]

    @property
    def ket(self):
        return self.args[1]

    def _eval_conjugate(self):
        return InnerProduct(Dagger(self.ket), Dagger(self.bra))

    def _sympyrepr(self, printer, *args):
        return '%s(%s,%s)' % (self.__class__.__name__,
            printer._print(self.bra, *args), printer._print(self.ket, *args))

    def _sympystr(self, printer, *args):
        sbra = printer._print(self.bra)
        sket = printer._print(self.ket)
        return '%s|%s' % (sbra[:-1], sket[1:])

    def _pretty(self, printer, *args):
        # Print state contents
        bra = self.bra._print_contents_pretty(printer, *args)
        ket = self.ket._print_contents_pretty(printer, *args)
        # Print brackets
        height = max(bra.height(), ket.height())
        use_unicode = printer._use_unicode
        lbracket, _ = self.bra._pretty_brackets(height, use_unicode)
        cbracket, rbracket = self.ket._pretty_brackets(height, use_unicode)
        # Build innerproduct
        pform = prettyForm(*bra.left(lbracket))
        pform = prettyForm(*pform.right(cbracket))
        pform = prettyForm(*pform.right(ket))
        pform = prettyForm(*pform.right(rbracket))
        return pform

    def _latex(self, printer, *args):
        bra_label = self.bra._print_contents_latex(printer, *args)
        ket = printer._print(self.ket, *args)
        return r'\left\langle %s \right. %s' % (bra_label, ket)

    def doit(self, **hints):
        try:
            r = self.ket._eval_innerproduct(self.bra, **hints)
        except NotImplementedError:
            try:
                r = conjugate(
                    self.bra.dual._eval_innerproduct(self.ket.dual, **hints)
                )
            except NotImplementedError:
                r = None
        if r is not None:
            return r
        return self

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\plotting\pygletplot\__init__.py ===
"""Plotting module that can plot 2D and 3D functions
"""

from sympy.utilities.decorator import doctest_depends_on

@doctest_depends_on(modules=('pyglet',))
def PygletPlot(*args, **kwargs):
    """

    Plot Examples
    =============

    See examples/advanced/pyglet_plotting.py for many more examples.

    >>> from sympy.plotting.pygletplot import PygletPlot as Plot
    >>> from sympy.abc import x, y, z

    >>> Plot(x*y**3-y*x**3)
    [0]: -x**3*y + x*y**3, 'mode=cartesian'

    >>> p = Plot()
    >>> p[1] = x*y
    >>> p[1].color = z, (0.4,0.4,0.9), (0.9,0.4,0.4)

    >>> p = Plot()
    >>> p[1] =  x**2+y**2
    >>> p[2] = -x**2-y**2


    Variable Intervals
    ==================

    The basic format is [var, min, max, steps], but the
    syntax is flexible and arguments left out are taken
    from the defaults for the current coordinate mode:

    >>> Plot(x**2) # implies [x,-5,5,100]
    [0]: x**2, 'mode=cartesian'

    >>> Plot(x**2, [], []) # [x,-1,1,40], [y,-1,1,40]
    [0]: x**2, 'mode=cartesian'
    >>> Plot(x**2-y**2, [100], [100]) # [x,-1,1,100], [y,-1,1,100]
    [0]: x**2 - y**2, 'mode=cartesian'
    >>> Plot(x**2, [x,-13,13,100])
    [0]: x**2, 'mode=cartesian'
    >>> Plot(x**2, [-13,13]) # [x,-13,13,100]
    [0]: x**2, 'mode=cartesian'
    >>> Plot(x**2, [x,-13,13]) # [x,-13,13,100]
    [0]: x**2, 'mode=cartesian'
    >>> Plot(1*x, [], [x], mode='cylindrical')
    ... # [unbound_theta,0,2*Pi,40], [x,-1,1,20]
    [0]: x, 'mode=cartesian'


    Coordinate Modes
    ================

    Plot supports several curvilinear coordinate modes, and
    they independent for each plotted function. You can specify
    a coordinate mode explicitly with the 'mode' named argument,
    but it can be automatically determined for Cartesian or
    parametric plots, and therefore must only be specified for
    polar, cylindrical, and spherical modes.

    Specifically, Plot(function arguments) and Plot[n] =
    (function arguments) will interpret your arguments as a
    Cartesian plot if you provide one function and a parametric
    plot if you provide two or three functions. Similarly, the
    arguments will be interpreted as a curve if one variable is
    used, and a surface if two are used.

    Supported mode names by number of variables:

    1: parametric, cartesian, polar
    2: parametric, cartesian, cylindrical = polar, spherical

    >>> Plot(1, mode='spherical')


    Calculator-like Interface
    =========================

    >>> p = Plot(visible=False)
    >>> f = x**2
    >>> p[1] = f
    >>> p[2] = f.diff(x)
    >>> p[3] = f.diff(x).diff(x)
    >>> p
    [1]: x**2, 'mode=cartesian'
    [2]: 2*x, 'mode=cartesian'
    [3]: 2, 'mode=cartesian'
    >>> p.show()
    >>> p.clear()
    >>> p
    <blank plot>
    >>> p[1] =  x**2+y**2
    >>> p[1].style = 'solid'
    >>> p[2] = -x**2-y**2
    >>> p[2].style = 'wireframe'
    >>> p[1].color = z, (0.4,0.4,0.9), (0.9,0.4,0.4)
    >>> p[1].style = 'both'
    >>> p[2].style = 'both'
    >>> p.close()


    Plot Window Keyboard Controls
    =============================

    Screen Rotation:
        X,Y axis      Arrow Keys, A,S,D,W, Numpad 4,6,8,2
        Z axis        Q,E, Numpad 7,9

    Model Rotation:
        Z axis        Z,C, Numpad 1,3

    Zoom:             R,F, PgUp,PgDn, Numpad +,-

    Reset Camera:     X, Numpad 5

    Camera Presets:
        XY            F1
        XZ            F2
        YZ            F3
        Perspective   F4

    Sensitivity Modifier: SHIFT

    Axes Toggle:
        Visible       F5
        Colors        F6

    Close Window:     ESCAPE

    =============================
    """

    from sympy.plotting.pygletplot.plot import PygletPlot
    return PygletPlot(*args, **kwargs)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\onnx_model_sam2.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

import logging

from fusion_attention_sam2 import FusionMultiHeadAttentionSam2
from fusion_layernorm import FusionLayerNormalizationNCHW
from fusion_options import FusionOptions
from import_utils import is_installed
from onnx import ModelProto
from onnx_model_bert import BertOnnxModel

logger = logging.getLogger(__name__)


class Sam2OnnxModel(BertOnnxModel):
    def __init__(self, model: ModelProto, num_heads: int = 0, hidden_size: int = 0):
        """Initialize SAM2 ONNX Model.

        Args:
            model (ModelProto): the ONNX model
            num_heads (int, optional): number of attention heads. Defaults to 0 (detect the parameter automatically).
            hidden_size (int, optional): hidden dimension. Defaults to 0 (detect the parameter automatically).
        """
        assert (num_heads == 0 and hidden_size == 0) or (num_heads > 0 and hidden_size % num_heads == 0)

        super().__init__(model, num_heads=num_heads, hidden_size=hidden_size)

    def postprocess(self):
        self.prune_graph()
        self.remove_unused_constant()

    def fuse_layer_norm(self):
        super().fuse_layer_norm()

        fusion = FusionLayerNormalizationNCHW(self)
        fusion.apply()

    def fuse_multi_head_attention(self, options: FusionOptions | None = None):
        mha_fusion = FusionMultiHeadAttentionSam2(self, self.hidden_size, self.num_heads)
        mha_fusion.apply()

    def optimize(self, options: FusionOptions | None = None, add_dynamic_axes: bool = False):
        if is_installed("tqdm"):
            import tqdm
            from tqdm.contrib.logging import logging_redirect_tqdm

            with logging_redirect_tqdm():
                steps = 12
                progress_bar = tqdm.tqdm(range(steps), initial=0, desc="sam2 fusion")
                self._optimize(options, progress_bar)
        else:
            logger.info("tqdm is not installed. Run optimization without progress bar")
            self._optimize(options, None)

    def _optimize(self, options: FusionOptions | None = None, progress_bar=None):
        if (options is not None) and not options.enable_shape_inference:
            self.disable_shape_inference()

        self.utils.remove_identity_nodes()
        if progress_bar:
            progress_bar.update(1)

        # Remove cast nodes that having same data type of input and output based on symbolic shape inference.
        self.utils.remove_useless_cast_nodes()
        if progress_bar:
            progress_bar.update(1)

        if (options is None) or options.enable_layer_norm:
            self.fuse_layer_norm()
        if progress_bar:
            progress_bar.update(1)

        if (options is None) or options.enable_gelu:
            self.fuse_gelu()
        if progress_bar:
            progress_bar.update(1)

        self.fuse_reshape()
        if progress_bar:
            progress_bar.update(1)

        if (options is None) or options.enable_attention:
            self.fuse_multi_head_attention(options)
        if progress_bar:
            progress_bar.update(1)

        if (options is None) or options.enable_skip_layer_norm:
            self.fuse_skip_layer_norm()
        if progress_bar:
            progress_bar.update(1)

        self.fuse_shape()
        if progress_bar:
            progress_bar.update(1)

        # Remove reshape nodes that having same shape of input and output based on symbolic shape inference.
        self.utils.remove_useless_reshape_nodes()
        if progress_bar:
            progress_bar.update(1)

        if (options is None) or options.enable_bias_skip_layer_norm:
            # Fuse SkipLayerNormalization and Add Bias before it.
            self.fuse_add_bias_skip_layer_norm()
        if progress_bar:
            progress_bar.update(1)

        if options is not None and options.enable_gelu_approximation:
            self.gelu_approximation()
        if progress_bar:
            progress_bar.update(1)

        self.postprocess()
        if progress_bar:
            progress_bar.update(1)

        logger.info(f"opset version: {self.get_opset_version()}")

    def get_fused_operator_statistics(self):
        """
        Returns node count of fused operators.
        """
        op_count = {}
        ops = [
            "MultiHeadAttention",
            "LayerNormalization",
            "SkipLayerNormalization",
        ]

        for op in ops:
            nodes = self.get_nodes_by_op_type(op)
            op_count[op] = len(nodes)

        logger.info(f"Optimized operators:{op_count}")
        return op_count

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\packaging\extended.py ===
# Copyright (c) 2010-2024 openpyxl


from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Typed,
)
from openpyxl.descriptors.nested import (
    NestedText,
)

from openpyxl.xml.constants import XPROPS_NS
from openpyxl import __version__


class DigSigBlob(Serialisable):

    __elements__ = __attrs__ = ()


class VectorLpstr(Serialisable):

    __elements__ = __attrs__ = ()


class VectorVariant(Serialisable):

    __elements__ = __attrs__ = ()


class ExtendedProperties(Serialisable):

    """
    See 22.2

    Most of this is irrelevant but Excel is very picky about the version number

    It uses XX.YYYY (Version.Build) and expects everyone else to

    We provide Major.Minor and the full version in the application name
    """

    tagname = "Properties"

    Template = NestedText(expected_type=str, allow_none=True)
    Manager = NestedText(expected_type=str, allow_none=True)
    Company = NestedText(expected_type=str, allow_none=True)
    Pages = NestedText(expected_type=int, allow_none=True)
    Words = NestedText(expected_type=int,allow_none=True)
    Characters = NestedText(expected_type=int, allow_none=True)
    PresentationFormat = NestedText(expected_type=str, allow_none=True)
    Lines = NestedText(expected_type=int, allow_none=True)
    Paragraphs = NestedText(expected_type=int, allow_none=True)
    Slides = NestedText(expected_type=int, allow_none=True)
    Notes = NestedText(expected_type=int, allow_none=True)
    TotalTime = NestedText(expected_type=int, allow_none=True)
    HiddenSlides = NestedText(expected_type=int, allow_none=True)
    MMClips = NestedText(expected_type=int, allow_none=True)
    ScaleCrop = NestedText(expected_type=bool, allow_none=True)
    HeadingPairs = Typed(expected_type=VectorVariant, allow_none=True)
    TitlesOfParts = Typed(expected_type=VectorLpstr, allow_none=True)
    LinksUpToDate = NestedText(expected_type=bool, allow_none=True)
    CharactersWithSpaces = NestedText(expected_type=int, allow_none=True)
    SharedDoc = NestedText(expected_type=bool, allow_none=True)
    HyperlinkBase = NestedText(expected_type=str, allow_none=True)
    HLinks = Typed(expected_type=VectorVariant, allow_none=True)
    HyperlinksChanged = NestedText(expected_type=bool, allow_none=True)
    DigSig = Typed(expected_type=DigSigBlob, allow_none=True)
    Application = NestedText(expected_type=str, allow_none=True)
    AppVersion = NestedText(expected_type=str, allow_none=True)
    DocSecurity = NestedText(expected_type=int, allow_none=True)

    __elements__ = ('Application', 'AppVersion', 'DocSecurity', 'ScaleCrop',
                    'LinksUpToDate', 'SharedDoc', 'HyperlinksChanged')

    def __init__(self,
                 Template=None,
                 Manager=None,
                 Company=None,
                 Pages=None,
                 Words=None,
                 Characters=None,
                 PresentationFormat=None,
                 Lines=None,
                 Paragraphs=None,
                 Slides=None,
                 Notes=None,
                 TotalTime=None,
                 HiddenSlides=None,
                 MMClips=None,
                 ScaleCrop=None,
                 HeadingPairs=None,
                 TitlesOfParts=None,
                 LinksUpToDate=None,
                 CharactersWithSpaces=None,
                 SharedDoc=None,
                 HyperlinkBase=None,
                 HLinks=None,
                 HyperlinksChanged=None,
                 DigSig=None,
                 Application=None,
                 AppVersion=None,
                 DocSecurity=None,
                ):
        self.Template = Template
        self.Manager = Manager
        self.Company = Company
        self.Pages = Pages
        self.Words = Words
        self.Characters = Characters
        self.PresentationFormat = PresentationFormat
        self.Lines = Lines
        self.Paragraphs = Paragraphs
        self.Slides = Slides
        self.Notes = Notes
        self.TotalTime = TotalTime
        self.HiddenSlides = HiddenSlides
        self.MMClips = MMClips
        self.ScaleCrop = ScaleCrop
        self.HeadingPairs = None
        self.TitlesOfParts = None
        self.LinksUpToDate = LinksUpToDate
        self.CharactersWithSpaces = CharactersWithSpaces
        self.SharedDoc = SharedDoc
        self.HyperlinkBase = HyperlinkBase
        self.HLinks = None
        self.HyperlinksChanged = HyperlinksChanged
        self.DigSig = None
        self.Application = f"Microsoft Excel Compatible / Openpyxl {__version__}"
        self.AppVersion = ".".join(__version__.split(".")[:-1])
        self.DocSecurity = DocSecurity


    def to_tree(self):
        tree = super().to_tree()
        tree.set("xmlns", XPROPS_NS)
        return tree

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\commands\__init__.py ===
"""
Package containing all pip commands
"""

import importlib
from collections import namedtuple
from typing import Any, Dict, Optional

from pip._internal.cli.base_command import Command

CommandInfo = namedtuple("CommandInfo", "module_path, class_name, summary")

# This dictionary does a bunch of heavy lifting for help output:
# - Enables avoiding additional (costly) imports for presenting `--help`.
# - The ordering matters for help display.
#
# Even though the module path starts with the same "pip._internal.commands"
# prefix, the full path makes testing easier (specifically when modifying
# `commands_dict` in test setup / teardown).
commands_dict: Dict[str, CommandInfo] = {
    "install": CommandInfo(
        "pip._internal.commands.install",
        "InstallCommand",
        "Install packages.",
    ),
    "lock": CommandInfo(
        "pip._internal.commands.lock",
        "LockCommand",
        "Generate a lock file.",
    ),
    "download": CommandInfo(
        "pip._internal.commands.download",
        "DownloadCommand",
        "Download packages.",
    ),
    "uninstall": CommandInfo(
        "pip._internal.commands.uninstall",
        "UninstallCommand",
        "Uninstall packages.",
    ),
    "freeze": CommandInfo(
        "pip._internal.commands.freeze",
        "FreezeCommand",
        "Output installed packages in requirements format.",
    ),
    "inspect": CommandInfo(
        "pip._internal.commands.inspect",
        "InspectCommand",
        "Inspect the python environment.",
    ),
    "list": CommandInfo(
        "pip._internal.commands.list",
        "ListCommand",
        "List installed packages.",
    ),
    "show": CommandInfo(
        "pip._internal.commands.show",
        "ShowCommand",
        "Show information about installed packages.",
    ),
    "check": CommandInfo(
        "pip._internal.commands.check",
        "CheckCommand",
        "Verify installed packages have compatible dependencies.",
    ),
    "config": CommandInfo(
        "pip._internal.commands.configuration",
        "ConfigurationCommand",
        "Manage local and global configuration.",
    ),
    "search": CommandInfo(
        "pip._internal.commands.search",
        "SearchCommand",
        "Search PyPI for packages.",
    ),
    "cache": CommandInfo(
        "pip._internal.commands.cache",
        "CacheCommand",
        "Inspect and manage pip's wheel cache.",
    ),
    "index": CommandInfo(
        "pip._internal.commands.index",
        "IndexCommand",
        "Inspect information available from package indexes.",
    ),
    "wheel": CommandInfo(
        "pip._internal.commands.wheel",
        "WheelCommand",
        "Build wheels from your requirements.",
    ),
    "hash": CommandInfo(
        "pip._internal.commands.hash",
        "HashCommand",
        "Compute hashes of package archives.",
    ),
    "completion": CommandInfo(
        "pip._internal.commands.completion",
        "CompletionCommand",
        "A helper command used for command completion.",
    ),
    "debug": CommandInfo(
        "pip._internal.commands.debug",
        "DebugCommand",
        "Show information useful for debugging.",
    ),
    "help": CommandInfo(
        "pip._internal.commands.help",
        "HelpCommand",
        "Show help for commands.",
    ),
}


def create_command(name: str, **kwargs: Any) -> Command:
    """
    Create an instance of the Command class with the given name.
    """
    module_path, class_name, summary = commands_dict[name]
    module = importlib.import_module(module_path)
    command_class = getattr(module, class_name)
    command = command_class(name=name, summary=summary, **kwargs)

    return command


def get_similar_commands(name: str) -> Optional[str]:
    """Command name auto-correct."""
    from difflib import get_close_matches

    name = name.lower()

    close_commands = get_close_matches(name, commands_dict.keys())

    if close_commands:
        return close_commands[0]
    else:
        return None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\urllib3\util\request.py ===
from __future__ import absolute_import

from base64 import b64encode

from ..exceptions import UnrewindableBodyError
from ..packages.six import b, integer_types

# Pass as a value within ``headers`` to skip
# emitting some HTTP headers that are added automatically.
# The only headers that are supported are ``Accept-Encoding``,
# ``Host``, and ``User-Agent``.
SKIP_HEADER = "@@@SKIP_HEADER@@@"
SKIPPABLE_HEADERS = frozenset(["accept-encoding", "host", "user-agent"])

ACCEPT_ENCODING = "gzip,deflate"

_FAILEDTELL = object()


def make_headers(
    keep_alive=None,
    accept_encoding=None,
    user_agent=None,
    basic_auth=None,
    proxy_basic_auth=None,
    disable_cache=None,
):
    """
    Shortcuts for generating request headers.

    :param keep_alive:
        If ``True``, adds 'connection: keep-alive' header.

    :param accept_encoding:
        Can be a boolean, list, or string.
        ``True`` translates to 'gzip,deflate'.
        List will get joined by comma.
        String will be used as provided.

    :param user_agent:
        String representing the user-agent you want, such as
        "python-urllib3/0.6"

    :param basic_auth:
        Colon-separated username:password string for 'authorization: basic ...'
        auth header.

    :param proxy_basic_auth:
        Colon-separated username:password string for 'proxy-authorization: basic ...'
        auth header.

    :param disable_cache:
        If ``True``, adds 'cache-control: no-cache' header.

    Example::

        >>> make_headers(keep_alive=True, user_agent="Batman/1.0")
        {'connection': 'keep-alive', 'user-agent': 'Batman/1.0'}
        >>> make_headers(accept_encoding=True)
        {'accept-encoding': 'gzip,deflate'}
    """
    headers = {}
    if accept_encoding:
        if isinstance(accept_encoding, str):
            pass
        elif isinstance(accept_encoding, list):
            accept_encoding = ",".join(accept_encoding)
        else:
            accept_encoding = ACCEPT_ENCODING
        headers["accept-encoding"] = accept_encoding

    if user_agent:
        headers["user-agent"] = user_agent

    if keep_alive:
        headers["connection"] = "keep-alive"

    if basic_auth:
        headers["authorization"] = "Basic " + b64encode(b(basic_auth)).decode("utf-8")

    if proxy_basic_auth:
        headers["proxy-authorization"] = "Basic " + b64encode(
            b(proxy_basic_auth)
        ).decode("utf-8")

    if disable_cache:
        headers["cache-control"] = "no-cache"

    return headers


def set_file_position(body, pos):
    """
    If a position is provided, move file to that point.
    Otherwise, we'll attempt to record a position for future use.
    """
    if pos is not None:
        rewind_body(body, pos)
    elif getattr(body, "tell", None) is not None:
        try:
            pos = body.tell()
        except (IOError, OSError):
            # This differentiates from None, allowing us to catch
            # a failed `tell()` later when trying to rewind the body.
            pos = _FAILEDTELL

    return pos


def rewind_body(body, body_pos):
    """
    Attempt to rewind body to a certain position.
    Primarily used for request redirects and retries.

    :param body:
        File-like object that supports seek.

    :param int pos:
        Position to seek to in file.
    """
    body_seek = getattr(body, "seek", None)
    if body_seek is not None and isinstance(body_pos, integer_types):
        try:
            body_seek(body_pos)
        except (IOError, OSError):
            raise UnrewindableBodyError(
                "An error occurred when rewinding request body for redirect/retry."
            )
    elif body_pos is _FAILEDTELL:
        raise UnrewindableBodyError(
            "Unable to record file position for rewinding "
            "request body during a redirect/retry."
        )
    else:
        raise ValueError(
            "body_pos must be of type integer, instead it was %s." % type(body_pos)
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\remote\shadowroot.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

from hashlib import md5 as md5_hash

from ..common.by import By
from .command import Command


class ShadowRoot:
    # TODO: We should look and see  how we can create a search context like Java/.NET

    def __init__(self, session, id_) -> None:
        self.session = session
        self._id = id_

    def __eq__(self, other_shadowroot) -> bool:
        return self._id == other_shadowroot._id

    def __hash__(self) -> int:
        return int(md5_hash(self._id.encode("utf-8")).hexdigest(), 16)

    def __repr__(self) -> str:
        return '<{0.__module__}.{0.__name__} (session="{1}", element="{2}")>'.format(
            type(self), self.session.session_id, self._id
        )

    @property
    def id(self) -> str:
        return self._id

    def find_element(self, by: str = By.ID, value: str = None):
        """Find an element inside a shadow root given a By strategy and
        locator.

        Parameters:
        -----------
        by : selenium.webdriver.common.by.By
            The locating strategy to use. Default is `By.ID`. Supported values include:
            - By.ID: Locate by element ID.
            - By.NAME: Locate by the `name` attribute.
            - By.XPATH: Locate by an XPath expression.
            - By.CSS_SELECTOR: Locate by a CSS selector.
            - By.CLASS_NAME: Locate by the `class` attribute.
            - By.TAG_NAME: Locate by the tag name (e.g., "input", "button").
            - By.LINK_TEXT: Locate a link element by its exact text.
            - By.PARTIAL_LINK_TEXT: Locate a link element by partial text match.
            - RelativeBy: Locate elements relative to a specified root element.

        Example:
        --------
        element = driver.find_element(By.ID, 'foo')

        Returns:
        -------
        WebElement
            The first matching `WebElement` found on the page.
        """
        if by == By.ID:
            by = By.CSS_SELECTOR
            value = f'[id="{value}"]'
        elif by == By.CLASS_NAME:
            by = By.CSS_SELECTOR
            value = f".{value}"
        elif by == By.NAME:
            by = By.CSS_SELECTOR
            value = f'[name="{value}"]'

        return self._execute(Command.FIND_ELEMENT_FROM_SHADOW_ROOT, {"using": by, "value": value})["value"]

    def find_elements(self, by: str = By.ID, value: str = None):
        """Find elements inside a shadow root given a By strategy and locator.

        Parameters:
        -----------
        by : selenium.webdriver.common.by.By
            The locating strategy to use. Default is `By.ID`. Supported values include:
            - By.ID: Locate by element ID.
            - By.NAME: Locate by the `name` attribute.
            - By.XPATH: Locate by an XPath expression.
            - By.CSS_SELECTOR: Locate by a CSS selector.
            - By.CLASS_NAME: Locate by the `class` attribute.
            - By.TAG_NAME: Locate by the tag name (e.g., "input", "button").
            - By.LINK_TEXT: Locate a link element by its exact text.
            - By.PARTIAL_LINK_TEXT: Locate a link element by partial text match.
            - RelativeBy: Locate elements relative to a specified root element.

        Example:
        --------
        element = driver.find_elements(By.ID, 'foo')

        Returns:
        -------
        List[WebElement]
            list of `WebElements` matching locator strategy found on the page.
        """
        if by == By.ID:
            by = By.CSS_SELECTOR
            value = f'[id="{value}"]'
        elif by == By.CLASS_NAME:
            by = By.CSS_SELECTOR
            value = f".{value}"
        elif by == By.NAME:
            by = By.CSS_SELECTOR
            value = f'[name="{value}"]'

        return self._execute(Command.FIND_ELEMENTS_FROM_SHADOW_ROOT, {"using": by, "value": value})["value"]

    # Private Methods
    def _execute(self, command, params=None):
        """Executes a command against the underlying HTML element.

        Args:
          command: The name of the command to _execute as a string.
          params: A dictionary of named parameters to send with the command.

        Returns:
          The command's JSON response loaded into a dictionary object.
        """
        if not params:
            params = {}
        params["shadowId"] = self._id
        return self.session.execute(command, params)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\urllib3\util\connection.py ===
from __future__ import annotations

import socket
import typing

from ..exceptions import LocationParseError
from .timeout import _DEFAULT_TIMEOUT, _TYPE_TIMEOUT

_TYPE_SOCKET_OPTIONS = list[tuple[int, int, typing.Union[int, bytes]]]

if typing.TYPE_CHECKING:
    from .._base_connection import BaseHTTPConnection


def is_connection_dropped(conn: BaseHTTPConnection) -> bool:  # Platform-specific
    """
    Returns True if the connection is dropped and should be closed.
    :param conn: :class:`urllib3.connection.HTTPConnection` object.
    """
    return not conn.is_connected


# This function is copied from socket.py in the Python 2.7 standard
# library test suite. Added to its signature is only `socket_options`.
# One additional modification is that we avoid binding to IPv6 servers
# discovered in DNS if the system doesn't have IPv6 functionality.
def create_connection(
    address: tuple[str, int],
    timeout: _TYPE_TIMEOUT = _DEFAULT_TIMEOUT,
    source_address: tuple[str, int] | None = None,
    socket_options: _TYPE_SOCKET_OPTIONS | None = None,
) -> socket.socket:
    """Connect to *address* and return the socket object.

    Convenience function.  Connect to *address* (a 2-tuple ``(host,
    port)``) and return the socket object.  Passing the optional
    *timeout* parameter will set the timeout on the socket instance
    before attempting to connect.  If no *timeout* is supplied, the
    global default timeout setting returned by :func:`socket.getdefaulttimeout`
    is used.  If *source_address* is set it must be a tuple of (host, port)
    for the socket to bind as a source address before making the connection.
    An host of '' or port 0 tells the OS to use the default.
    """

    host, port = address
    if host.startswith("["):
        host = host.strip("[]")
    err = None

    # Using the value from allowed_gai_family() in the context of getaddrinfo lets
    # us select whether to work with IPv4 DNS records, IPv6 records, or both.
    # The original create_connection function always returns all records.
    family = allowed_gai_family()

    try:
        host.encode("idna")
    except UnicodeError:
        raise LocationParseError(f"'{host}', label empty or too long") from None

    for res in socket.getaddrinfo(host, port, family, socket.SOCK_STREAM):
        af, socktype, proto, canonname, sa = res
        sock = None
        try:
            sock = socket.socket(af, socktype, proto)

            # If provided, set socket level options before connecting.
            _set_socket_options(sock, socket_options)

            if timeout is not _DEFAULT_TIMEOUT:
                sock.settimeout(timeout)
            if source_address:
                sock.bind(source_address)
            sock.connect(sa)
            # Break explicitly a reference cycle
            err = None
            return sock

        except OSError as _:
            err = _
            if sock is not None:
                sock.close()

    if err is not None:
        try:
            raise err
        finally:
            # Break explicitly a reference cycle
            err = None
    else:
        raise OSError("getaddrinfo returns an empty list")


def _set_socket_options(
    sock: socket.socket, options: _TYPE_SOCKET_OPTIONS | None
) -> None:
    if options is None:
        return

    for opt in options:
        sock.setsockopt(*opt)


def allowed_gai_family() -> socket.AddressFamily:
    """This function is designed to work in the context of
    getaddrinfo, where family=socket.AF_UNSPEC is the default and
    will perform a DNS search for both IPv6 and IPv4 records."""

    family = socket.AF_INET
    if HAS_IPV6:
        family = socket.AF_UNSPEC
    return family


def _has_ipv6(host: str) -> bool:
    """Returns True if the system can bind an IPv6 address."""
    sock = None
    has_ipv6 = False

    if socket.has_ipv6:
        # has_ipv6 returns true if cPython was compiled with IPv6 support.
        # It does not tell us if the system has IPv6 support enabled. To
        # determine that we must bind to an IPv6 address.
        # https://github.com/urllib3/urllib3/pull/611
        # https://bugs.python.org/issue658327
        try:
            sock = socket.socket(socket.AF_INET6)
            sock.bind((host, 0))
            has_ipv6 = True
        except Exception:
            pass

    if sock:
        sock.close()
    return has_ipv6


HAS_IPV6 = _has_ipv6("::1")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\io\test_sql.py ===
from __future__ import annotations

import contextlib
from contextlib import closing
import csv
from datetime import (
    date,
    datetime,
    time,
    timedelta,
)
from io import StringIO
from pathlib import Path
import sqlite3
from typing import TYPE_CHECKING
import uuid

import numpy as np
import pytest

from pandas._config import using_string_dtype

from pandas._libs import lib
from pandas.compat import (
    pa_version_under13p0,
    pa_version_under14p1,
)
from pandas.compat._optional import import_optional_dependency
import pandas.util._test_decorators as td

import pandas as pd
from pandas import (
    DataFrame,
    Index,
    MultiIndex,
    Series,
    Timestamp,
    concat,
    date_range,
    isna,
    to_datetime,
    to_timedelta,
)
import pandas._testing as tm
from pandas.util.version import Version

from pandas.io import sql
from pandas.io.sql import (
    SQLAlchemyEngine,
    SQLDatabase,
    SQLiteDatabase,
    get_engine,
    pandasSQL_builder,
    read_sql_query,
    read_sql_table,
)

if TYPE_CHECKING:
    import sqlalchemy


pytestmark = [
    pytest.mark.filterwarnings(
        "ignore:Passing a BlockManager to DataFrame:DeprecationWarning"
    ),
    pytest.mark.single_cpu,
]


@pytest.fixture
def sql_strings():
    return {
        "read_parameters": {
            "sqlite": "SELECT * FROM iris WHERE Name=? AND SepalLength=?",
            "mysql": "SELECT * FROM iris WHERE `Name`=%s AND `SepalLength`=%s",
            "postgresql": 'SELECT * FROM iris WHERE "Name"=%s AND "SepalLength"=%s',
        },
        "read_named_parameters": {
            "sqlite": """
                SELECT * FROM iris WHERE Name=:name AND SepalLength=:length
                """,
            "mysql": """
                SELECT * FROM iris WHERE
                `Name`=%(name)s AND `SepalLength`=%(length)s
                """,
            "postgresql": """
                SELECT * FROM iris WHERE
                "Name"=%(name)s AND "SepalLength"=%(length)s
                """,
        },
        "read_no_parameters_with_percent": {
            "sqlite": "SELECT * FROM iris WHERE Name LIKE '%'",
            "mysql": "SELECT * FROM iris WHERE `Name` LIKE '%'",
            "postgresql": "SELECT * FROM iris WHERE \"Name\" LIKE '%'",
        },
    }


def iris_table_metadata():
    import sqlalchemy
    from sqlalchemy import (
        Column,
        Double,
        Float,
        MetaData,
        String,
        Table,
    )

    dtype = Double if Version(sqlalchemy.__version__) >= Version("2.0.0") else Float
    metadata = MetaData()
    iris = Table(
        "iris",
        metadata,
        Column("SepalLength", dtype),
        Column("SepalWidth", dtype),
        Column("PetalLength", dtype),
        Column("PetalWidth", dtype),
        Column("Name", String(200)),
    )
    return iris


def create_and_load_iris_sqlite3(conn, iris_file: Path):
    stmt = """CREATE TABLE iris (
            "SepalLength" REAL,
            "SepalWidth" REAL,
            "PetalLength" REAL,
            "PetalWidth" REAL,
            "Name" TEXT
        )"""

    cur = conn.cursor()
    cur.execute(stmt)
    with iris_file.open(newline=None, encoding="utf-8") as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        stmt = "INSERT INTO iris VALUES(?, ?, ?, ?, ?)"
        # ADBC requires explicit types - no implicit str -> float conversion
        records = []
        records = [
            (
                float(row[0]),
                float(row[1]),
                float(row[2]),
                float(row[3]),
                row[4],
            )
            for row in reader
        ]

        cur.executemany(stmt, records)
    cur.close()

    conn.commit()


def create_and_load_iris_postgresql(conn, iris_file: Path):
    stmt = """CREATE TABLE iris (
            "SepalLength" DOUBLE PRECISION,
            "SepalWidth" DOUBLE PRECISION,
            "PetalLength" DOUBLE PRECISION,
            "PetalWidth" DOUBLE PRECISION,
            "Name" TEXT
        )"""
    with conn.cursor() as cur:
        cur.execute(stmt)
        with iris_file.open(newline=None, encoding="utf-8") as csvfile:
            reader = csv.reader(csvfile)
            next(reader)
            stmt = "INSERT INTO iris VALUES($1, $2, $3, $4, $5)"
            # ADBC requires explicit types - no implicit str -> float conversion
            records = [
                (
                    float(row[0]),
                    float(row[1]),
                    float(row[2]),
                    float(row[3]),
                    row[4],
                )
                for row in reader
            ]

            cur.executemany(stmt, records)

    conn.commit()


def create_and_load_iris(conn, iris_file: Path):
    from sqlalchemy import insert

    iris = iris_table_metadata()

    with iris_file.open(newline=None, encoding="utf-8") as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader)
        params = [dict(zip(header, row)) for row in reader]
        stmt = insert(iris).values(params)
        with conn.begin() as con:
            iris.drop(con, checkfirst=True)
            iris.create(bind=con)
            con.execute(stmt)


def create_and_load_iris_view(conn):
    stmt = "CREATE VIEW iris_view AS SELECT * FROM iris"
    if isinstance(conn, sqlite3.Connection):
        cur = conn.cursor()
        cur.execute(stmt)
    else:
        adbc = import_optional_dependency("adbc_driver_manager.dbapi", errors="ignore")
        if adbc and isinstance(conn, adbc.Connection):
            with conn.cursor() as cur:
                cur.execute(stmt)
            conn.commit()
        else:
            from sqlalchemy import text

            stmt = text(stmt)
            with conn.begin() as con:
                con.execute(stmt)


def types_table_metadata(dialect: str):
    from sqlalchemy import (
        TEXT,
        Boolean,
        Column,
        DateTime,
        Float,
        Integer,
        MetaData,
        Table,
    )

    date_type = TEXT if dialect == "sqlite" else DateTime
    bool_type = Integer if dialect == "sqlite" else Boolean
    metadata = MetaData()
    types = Table(
        "types",
        metadata,
        Column("TextCol", TEXT),
        Column("DateCol", date_type),
        Column("IntDateCol", Integer),
        Column("IntDateOnlyCol", Integer),
        Column("FloatCol", Float),
        Column("IntCol", Integer),
        Column("BoolCol", bool_type),
        Column("IntColWithNull", Integer),
        Column("BoolColWithNull", bool_type),
    )
    return types


def create_and_load_types_sqlite3(conn, types_data: list[dict]):
    stmt = """CREATE TABLE types (
                    "TextCol" TEXT,
                    "DateCol" TEXT,
                    "IntDateCol" INTEGER,
                    "IntDateOnlyCol" INTEGER,
                    "FloatCol" REAL,
                    "IntCol" INTEGER,
                    "BoolCol" INTEGER,
                    "IntColWithNull" INTEGER,
                    "BoolColWithNull" INTEGER
                )"""

    ins_stmt = """
                INSERT INTO types
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
                """

    if isinstance(conn, sqlite3.Connection):
        cur = conn.cursor()
        cur.execute(stmt)
        cur.executemany(ins_stmt, types_data)
    else:
        with conn.cursor() as cur:
            cur.execute(stmt)
            cur.executemany(ins_stmt, types_data)

        conn.commit()


def create_and_load_types_postgresql(conn, types_data: list[dict]):
    with conn.cursor() as cur:
        stmt = """CREATE TABLE types (
                        "TextCol" TEXT,
                        "DateCol" TIMESTAMP,
                        "IntDateCol" INTEGER,
                        "IntDateOnlyCol" INTEGER,
                        "FloatCol" DOUBLE PRECISION,
                        "IntCol" INTEGER,
                        "BoolCol" BOOLEAN,
                        "IntColWithNull" INTEGER,
                        "BoolColWithNull" BOOLEAN
                    )"""
        cur.execute(stmt)

        stmt = """
                INSERT INTO types
                VALUES($1, $2::timestamp, $3, $4, $5, $6, $7, $8, $9)
                """

        cur.executemany(stmt, types_data)

    conn.commit()


def create_and_load_types(conn, types_data: list[dict], dialect: str):
    from sqlalchemy import insert
    from sqlalchemy.engine import Engine

    types = types_table_metadata(dialect)

    stmt = insert(types).values(types_data)
    if isinstance(conn, Engine):
        with conn.connect() as conn:
            with conn.begin():
                types.drop(conn, checkfirst=True)
                types.create(bind=conn)
                conn.execute(stmt)
    else:
        with conn.begin():
            types.drop(conn, checkfirst=True)
            types.create(bind=conn)
            conn.execute(stmt)


def create_and_load_postgres_datetz(conn):
    from sqlalchemy import (
        Column,
        DateTime,
        MetaData,
        Table,
        insert,
    )
    from sqlalchemy.engine import Engine

    metadata = MetaData()
    datetz = Table("datetz", metadata, Column("DateColWithTz", DateTime(timezone=True)))
    datetz_data = [
        {
            "DateColWithTz": "2000-01-01 00:00:00-08:00",
        },
        {
            "DateColWithTz": "2000-06-01 00:00:00-07:00",
        },
    ]
    stmt = insert(datetz).values(datetz_data)
    if isinstance(conn, Engine):
        with conn.connect() as conn:
            with conn.begin():
                datetz.drop(conn, checkfirst=True)
                datetz.create(bind=conn)
                conn.execute(stmt)
    else:
        with conn.begin():
            datetz.drop(conn, checkfirst=True)
            datetz.create(bind=conn)
            conn.execute(stmt)

    # "2000-01-01 00:00:00-08:00" should convert to
    # "2000-01-01 08:00:00"
    # "2000-06-01 00:00:00-07:00" should convert to
    # "2000-06-01 07:00:00"
    # GH 6415
    expected_data = [
        Timestamp("2000-01-01 08:00:00", tz="UTC"),
        Timestamp("2000-06-01 07:00:00", tz="UTC"),
    ]
    return Series(expected_data, name="DateColWithTz")


def check_iris_frame(frame: DataFrame):
    pytype = frame.dtypes.iloc[0].type
    row = frame.iloc[0]
    assert issubclass(pytype, np.floating)
    tm.assert_series_equal(
        row, Series([5.1, 3.5, 1.4, 0.2, "Iris-setosa"], index=frame.columns, name=0)
    )
    assert frame.shape in ((150, 5), (8, 5))


def count_rows(conn, table_name: str):
    stmt = f"SELECT count(*) AS count_1 FROM {table_name}"
    adbc = import_optional_dependency("adbc_driver_manager.dbapi", errors="ignore")
    if isinstance(conn, sqlite3.Connection):
        cur = conn.cursor()
        return cur.execute(stmt).fetchone()[0]
    elif adbc and isinstance(conn, adbc.Connection):
        with conn.cursor() as cur:
            cur.execute(stmt)
            return cur.fetchone()[0]
    else:
        from sqlalchemy import create_engine
        from sqlalchemy.engine import Engine

        if isinstance(conn, str):
            try:
                engine = create_engine(conn)
                with engine.connect() as conn:
                    return conn.exec_driver_sql(stmt).scalar_one()
            finally:
                engine.dispose()
        elif isinstance(conn, Engine):
            with conn.connect() as conn:
                return conn.exec_driver_sql(stmt).scalar_one()
        else:
            return conn.exec_driver_sql(stmt).scalar_one()


@pytest.fixture
def iris_path(datapath):
    iris_path = datapath("io", "data", "csv", "iris.csv")
    return Path(iris_path)


@pytest.fixture
def types_data():
    return [
        {
            "TextCol": "first",
            "DateCol": "2000-01-03 00:00:00",
            "IntDateCol": 535852800,
            "IntDateOnlyCol": 20101010,
            "FloatCol": 10.10,
            "IntCol": 1,
            "BoolCol": False,
            "IntColWithNull": 1,
            "BoolColWithNull": False,
        },
        {
            "TextCol": "first",
            "DateCol": "2000-01-04 00:00:00",
            "IntDateCol": 1356998400,
            "IntDateOnlyCol": 20101212,
            "FloatCol": 10.10,
            "IntCol": 1,
            "BoolCol": False,
            "IntColWithNull": None,
            "BoolColWithNull": None,
        },
    ]


@pytest.fixture
def types_data_frame(types_data):
    dtypes = {
        "TextCol": "str",
        "DateCol": "str",
        "IntDateCol": "int64",
        "IntDateOnlyCol": "int64",
        "FloatCol": "float",
        "IntCol": "int64",
        "BoolCol": "int64",
        "IntColWithNull": "float",
        "BoolColWithNull": "float",
    }
    df = DataFrame(types_data)
    return df[dtypes.keys()].astype(dtypes)


@pytest.fixture
def test_frame1():
    columns = ["index", "A", "B", "C", "D"]
    data = [
        (
            "2000-01-03 00:00:00",
            0.980268513777,
            3.68573087906,
            -0.364216805298,
            -1.15973806169,
        ),
        (
            "2000-01-04 00:00:00",
            1.04791624281,
            -0.0412318367011,
            -0.16181208307,
            0.212549316967,
        ),
        (
            "2000-01-05 00:00:00",
            0.498580885705,
            0.731167677815,
            -0.537677223318,
            1.34627041952,
        ),
        (
            "2000-01-06 00:00:00",
            1.12020151869,
            1.56762092543,
            0.00364077397681,
            0.67525259227,
        ),
    ]
    return DataFrame(data, columns=columns)


@pytest.fixture
def test_frame3():
    columns = ["index", "A", "B"]
    data = [
        ("2000-01-03 00:00:00", 2**31 - 1, -1.987670),
        ("2000-01-04 00:00:00", -29, -0.0412318367011),
        ("2000-01-05 00:00:00", 20000, 0.731167677815),
        ("2000-01-06 00:00:00", -290867, 1.56762092543),
    ]
    return DataFrame(data, columns=columns)


def get_all_views(conn):
    if isinstance(conn, sqlite3.Connection):
        c = conn.execute("SELECT name FROM sqlite_master WHERE type='view'")
        return [view[0] for view in c.fetchall()]
    else:
        adbc = import_optional_dependency("adbc_driver_manager.dbapi", errors="ignore")
        if adbc and isinstance(conn, adbc.Connection):
            results = []
            info = conn.adbc_get_objects().read_all().to_pylist()
            for catalog in info:
                catalog["catalog_name"]
                for schema in catalog["catalog_db_schemas"]:
                    schema["db_schema_name"]
                    for table in schema["db_schema_tables"]:
                        if table["table_type"] == "view":
                            view_name = table["table_name"]
                            results.append(view_name)

            return results
        else:
            from sqlalchemy import inspect

            return inspect(conn).get_view_names()


def get_all_tables(conn):
    if isinstance(conn, sqlite3.Connection):
        c = conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
        return [table[0] for table in c.fetchall()]
    else:
        adbc = import_optional_dependency("adbc_driver_manager.dbapi", errors="ignore")

        if adbc and isinstance(conn, adbc.Connection):
            results = []
            info = conn.adbc_get_objects().read_all().to_pylist()
            for catalog in info:
                for schema in catalog["catalog_db_schemas"]:
                    for table in schema["db_schema_tables"]:
                        if table["table_type"] == "table":
                            table_name = table["table_name"]
                            results.append(table_name)

            return results
        else:
            from sqlalchemy import inspect

            return inspect(conn).get_table_names()


def drop_table(
    table_name: str,
    conn: sqlite3.Connection | sqlalchemy.engine.Engine | sqlalchemy.engine.Connection,
):
    if isinstance(conn, sqlite3.Connection):
        conn.execute(f"DROP TABLE IF EXISTS {sql._get_valid_sqlite_name(table_name)}")
        conn.commit()

    else:
        adbc = import_optional_dependency("adbc_driver_manager.dbapi", errors="ignore")
        if adbc and isinstance(conn, adbc.Connection):
            with conn.cursor() as cur:
                cur.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        else:
            with conn.begin() as con:
                with sql.SQLDatabase(con) as db:
                    db.drop_table(table_name)


def drop_view(
    view_name: str,
    conn: sqlite3.Connection | sqlalchemy.engine.Engine | sqlalchemy.engine.Connection,
):
    import sqlalchemy

    if isinstance(conn, sqlite3.Connection):
        conn.execute(f"DROP VIEW IF EXISTS {sql._get_valid_sqlite_name(view_name)}")
        conn.commit()
    else:
        adbc = import_optional_dependency("adbc_driver_manager.dbapi", errors="ignore")
        if adbc and isinstance(conn, adbc.Connection):
            with conn.cursor() as cur:
                cur.execute(f'DROP VIEW IF EXISTS "{view_name}"')
        else:
            quoted_view = conn.engine.dialect.identifier_preparer.quote_identifier(
                view_name
            )
            stmt = sqlalchemy.text(f"DROP VIEW IF EXISTS {quoted_view}")
            with conn.begin() as con:
                con.execute(stmt)  # type: ignore[union-attr]


@pytest.fixture
def mysql_pymysql_engine():
    sqlalchemy = pytest.importorskip("sqlalchemy")
    pymysql = pytest.importorskip("pymysql")
    engine = sqlalchemy.create_engine(
        "mysql+pymysql://root@localhost:3306/pandas",
        connect_args={"client_flag": pymysql.constants.CLIENT.MULTI_STATEMENTS},
        poolclass=sqlalchemy.pool.NullPool,
    )
    yield engine
    for view in get_all_views(engine):
        drop_view(view, engine)
    for tbl in get_all_tables(engine):
        drop_table(tbl, engine)
    engine.dispose()


@pytest.fixture
def mysql_pymysql_engine_iris(mysql_pymysql_engine, iris_path):
    create_and_load_iris(mysql_pymysql_engine, iris_path)
    create_and_load_iris_view(mysql_pymysql_engine)
    yield mysql_pymysql_engine


@pytest.fixture
def mysql_pymysql_engine_types(mysql_pymysql_engine, types_data):
    create_and_load_types(mysql_pymysql_engine, types_data, "mysql")
    yield mysql_pymysql_engine


@pytest.fixture
def mysql_pymysql_conn(mysql_pymysql_engine):
    with mysql_pymysql_engine.connect() as conn:
        yield conn


@pytest.fixture
def mysql_pymysql_conn_iris(mysql_pymysql_engine_iris):
    with mysql_pymysql_engine_iris.connect() as conn:
        yield conn


@pytest.fixture
def mysql_pymysql_conn_types(mysql_pymysql_engine_types):
    with mysql_pymysql_engine_types.connect() as conn:
        yield conn


@pytest.fixture
def postgresql_psycopg2_engine():
    sqlalchemy = pytest.importorskip("sqlalchemy")
    pytest.importorskip("psycopg2")
    engine = sqlalchemy.create_engine(
        "postgresql+psycopg2://postgres:postgres@localhost:5432/pandas",
        poolclass=sqlalchemy.pool.NullPool,
    )
    yield engine
    for view in get_all_views(engine):
        drop_view(view, engine)
    for tbl in get_all_tables(engine):
        drop_table(tbl, engine)
    engine.dispose()


@pytest.fixture
def postgresql_psycopg2_engine_iris(postgresql_psycopg2_engine, iris_path):
    create_and_load_iris(postgresql_psycopg2_engine, iris_path)
    create_and_load_iris_view(postgresql_psycopg2_engine)
    yield postgresql_psycopg2_engine


@pytest.fixture
def postgresql_psycopg2_engine_types(postgresql_psycopg2_engine, types_data):
    create_and_load_types(postgresql_psycopg2_engine, types_data, "postgres")
    yield postgresql_psycopg2_engine


@pytest.fixture
def postgresql_psycopg2_conn(postgresql_psycopg2_engine):
    with postgresql_psycopg2_engine.connect() as conn:
        yield conn


@pytest.fixture
def postgresql_adbc_conn():
    pytest.importorskip("pyarrow")
    pytest.importorskip("adbc_driver_postgresql")
    from adbc_driver_postgresql import dbapi

    uri = "postgresql://postgres:postgres@localhost:5432/pandas"
    with dbapi.connect(uri) as conn:
        yield conn
        for view in get_all_views(conn):
            drop_view(view, conn)
        for tbl in get_all_tables(conn):
            drop_table(tbl, conn)
        conn.commit()


@pytest.fixture
def postgresql_adbc_iris(postgresql_adbc_conn, iris_path):
    import adbc_driver_manager as mgr

    conn = postgresql_adbc_conn

    try:
        conn.adbc_get_table_schema("iris")
    except mgr.ProgrammingError:
        conn.rollback()
        create_and_load_iris_postgresql(conn, iris_path)
    try:
        conn.adbc_get_table_schema("iris_view")
    except mgr.ProgrammingError:  # note arrow-adbc issue 1022
        conn.rollback()
        create_and_load_iris_view(conn)
    yield conn


@pytest.fixture
def postgresql_adbc_types(postgresql_adbc_conn, types_data):
    import adbc_driver_manager as mgr

    conn = postgresql_adbc_conn

    try:
        conn.adbc_get_table_schema("types")
    except mgr.ProgrammingError:
        conn.rollback()
        new_data = [tuple(entry.values()) for entry in types_data]

        create_and_load_types_postgresql(conn, new_data)

    yield conn


@pytest.fixture
def postgresql_psycopg2_conn_iris(postgresql_psycopg2_engine_iris):
    with postgresql_psycopg2_engine_iris.connect() as conn:
        yield conn


@pytest.fixture
def postgresql_psycopg2_conn_types(postgresql_psycopg2_engine_types):
    with postgresql_psycopg2_engine_types.connect() as conn:
        yield conn


@pytest.fixture
def sqlite_str():
    pytest.importorskip("sqlalchemy")
    with tm.ensure_clean() as name:
        yield f"sqlite:///{name}"


@pytest.fixture
def sqlite_engine(sqlite_str):
    sqlalchemy = pytest.importorskip("sqlalchemy")
    engine = sqlalchemy.create_engine(sqlite_str, poolclass=sqlalchemy.pool.NullPool)
    yield engine
    for view in get_all_views(engine):
        drop_view(view, engine)
    for tbl in get_all_tables(engine):
        drop_table(tbl, engine)
    engine.dispose()


@pytest.fixture
def sqlite_conn(sqlite_engine):
    with sqlite_engine.connect() as conn:
        yield conn


@pytest.fixture
def sqlite_str_iris(sqlite_str, iris_path):
    sqlalchemy = pytest.importorskip("sqlalchemy")
    engine = sqlalchemy.create_engine(sqlite_str)
    create_and_load_iris(engine, iris_path)
    create_and_load_iris_view(engine)
    engine.dispose()
    return sqlite_str


@pytest.fixture
def sqlite_engine_iris(sqlite_engine, iris_path):
    create_and_load_iris(sqlite_engine, iris_path)
    create_and_load_iris_view(sqlite_engine)
    yield sqlite_engine


@pytest.fixture
def sqlite_conn_iris(sqlite_engine_iris):
    with sqlite_engine_iris.connect() as conn:
        yield conn


@pytest.fixture
def sqlite_str_types(sqlite_str, types_data):
    sqlalchemy = pytest.importorskip("sqlalchemy")
    engine = sqlalchemy.create_engine(sqlite_str)
    create_and_load_types(engine, types_data, "sqlite")
    engine.dispose()
    return sqlite_str


@pytest.fixture
def sqlite_engine_types(sqlite_engine, types_data):
    create_and_load_types(sqlite_engine, types_data, "sqlite")
    yield sqlite_engine


@pytest.fixture
def sqlite_conn_types(sqlite_engine_types):
    with sqlite_engine_types.connect() as conn:
        yield conn


@pytest.fixture
def sqlite_adbc_conn():
    pytest.importorskip("pyarrow")
    pytest.importorskip("adbc_driver_sqlite")
    from adbc_driver_sqlite import dbapi

    with tm.ensure_clean() as name:
        uri = f"file:{name}"
        with dbapi.connect(uri) as conn:
            yield conn
            for view in get_all_views(conn):
                drop_view(view, conn)
            for tbl in get_all_tables(conn):
                drop_table(tbl, conn)
            conn.commit()


@pytest.fixture
def sqlite_adbc_iris(sqlite_adbc_conn, iris_path):
    import adbc_driver_manager as mgr

    conn = sqlite_adbc_conn
    try:
        conn.adbc_get_table_schema("iris")
    except mgr.ProgrammingError:
        conn.rollback()
        create_and_load_iris_sqlite3(conn, iris_path)
    try:
        conn.adbc_get_table_schema("iris_view")
    except mgr.ProgrammingError:
        conn.rollback()
        create_and_load_iris_view(conn)
    yield conn


@pytest.fixture
def sqlite_adbc_types(sqlite_adbc_conn, types_data):
    import adbc_driver_manager as mgr

    conn = sqlite_adbc_conn
    try:
        conn.adbc_get_table_schema("types")
    except mgr.ProgrammingError:
        conn.rollback()
        new_data = []
        for entry in types_data:
            entry["BoolCol"] = int(entry["BoolCol"])
            if entry["BoolColWithNull"] is not None:
                entry["BoolColWithNull"] = int(entry["BoolColWithNull"])
            new_data.append(tuple(entry.values()))

        create_and_load_types_sqlite3(conn, new_data)
        conn.commit()

    yield conn


@pytest.fixture
def sqlite_buildin():
    with contextlib.closing(sqlite3.connect(":memory:")) as closing_conn:
        with closing_conn as conn:
            yield conn


@pytest.fixture
def sqlite_buildin_iris(sqlite_buildin, iris_path):
    create_and_load_iris_sqlite3(sqlite_buildin, iris_path)
    create_and_load_iris_view(sqlite_buildin)
    yield sqlite_buildin


@pytest.fixture
def sqlite_buildin_types(sqlite_buildin, types_data):
    types_data = [tuple(entry.values()) for entry in types_data]
    create_and_load_types_sqlite3(sqlite_buildin, types_data)
    yield sqlite_buildin


mysql_connectable = [
    pytest.param("mysql_pymysql_engine", marks=pytest.mark.db),
    pytest.param("mysql_pymysql_conn", marks=pytest.mark.db),
]

mysql_connectable_iris = [
    pytest.param("mysql_pymysql_engine_iris", marks=pytest.mark.db),
    pytest.param("mysql_pymysql_conn_iris", marks=pytest.mark.db),
]

mysql_connectable_types = [
    pytest.param("mysql_pymysql_engine_types", marks=pytest.mark.db),
    pytest.param("mysql_pymysql_conn_types", marks=pytest.mark.db),
]

postgresql_connectable = [
    pytest.param("postgresql_psycopg2_engine", marks=pytest.mark.db),
    pytest.param("postgresql_psycopg2_conn", marks=pytest.mark.db),
]

postgresql_connectable_iris = [
    pytest.param("postgresql_psycopg2_engine_iris", marks=pytest.mark.db),
    pytest.param("postgresql_psycopg2_conn_iris", marks=pytest.mark.db),
]

postgresql_connectable_types = [
    pytest.param("postgresql_psycopg2_engine_types", marks=pytest.mark.db),
    pytest.param("postgresql_psycopg2_conn_types", marks=pytest.mark.db),
]

sqlite_connectable = [
    "sqlite_engine",
    "sqlite_conn",
    "sqlite_str",
]

sqlite_connectable_iris = [
    "sqlite_engine_iris",
    "sqlite_conn_iris",
    "sqlite_str_iris",
]

sqlite_connectable_types = [
    "sqlite_engine_types",
    "sqlite_conn_types",
    "sqlite_str_types",
]

sqlalchemy_connectable = mysql_connectable + postgresql_connectable + sqlite_connectable

sqlalchemy_connectable_iris = (
    mysql_connectable_iris + postgresql_connectable_iris + sqlite_connectable_iris
)

sqlalchemy_connectable_types = (
    mysql_connectable_types + postgresql_connectable_types + sqlite_connectable_types
)

adbc_connectable = [
    "sqlite_adbc_conn",
    pytest.param("postgresql_adbc_conn", marks=pytest.mark.db),
]

adbc_connectable_iris = [
    pytest.param("postgresql_adbc_iris", marks=pytest.mark.db),
    "sqlite_adbc_iris",
]

adbc_connectable_types = [
    pytest.param("postgresql_adbc_types", marks=pytest.mark.db),
    "sqlite_adbc_types",
]


all_connectable = sqlalchemy_connectable + ["sqlite_buildin"] + adbc_connectable

all_connectable_iris = (
    sqlalchemy_connectable_iris + ["sqlite_buildin_iris"] + adbc_connectable_iris
)

all_connectable_types = (
    sqlalchemy_connectable_types + ["sqlite_buildin_types"] + adbc_connectable_types
)


@pytest.mark.parametrize("conn", all_connectable)
def test_dataframe_to_sql(conn, test_frame1, request):
    # GH 51086 if conn is sqlite_engine
    conn = request.getfixturevalue(conn)
    test_frame1.to_sql(name="test", con=conn, if_exists="append", index=False)


@pytest.mark.parametrize("conn", all_connectable)
def test_dataframe_to_sql_empty(conn, test_frame1, request):
    if conn == "postgresql_adbc_conn" and not using_string_dtype():
        request.node.add_marker(
            pytest.mark.xfail(
                reason="postgres ADBC driver < 1.2 cannot insert index with null type",
            )
        )

    # GH 51086 if conn is sqlite_engine
    conn = request.getfixturevalue(conn)
    empty_df = test_frame1.iloc[:0]
    empty_df.to_sql(name="test", con=conn, if_exists="append", index=False)


@pytest.mark.parametrize("conn", all_connectable)
def test_dataframe_to_sql_arrow_dtypes(conn, request):
    # GH 52046
    pytest.importorskip("pyarrow")
    df = DataFrame(
        {
            "int": pd.array([1], dtype="int8[pyarrow]"),
            "datetime": pd.array(
                [datetime(2023, 1, 1)], dtype="timestamp[ns][pyarrow]"
            ),
            "date": pd.array([date(2023, 1, 1)], dtype="date32[day][pyarrow]"),
            "timedelta": pd.array([timedelta(1)], dtype="duration[ns][pyarrow]"),
            "string": pd.array(["a"], dtype="string[pyarrow]"),
        }
    )

    if "adbc" in conn:
        if conn == "sqlite_adbc_conn":
            df = df.drop(columns=["timedelta"])
        if pa_version_under14p1:
            exp_warning = DeprecationWarning
            msg = "is_sparse is deprecated"
        else:
            exp_warning = None
            msg = ""
    else:
        exp_warning = UserWarning
        msg = "the 'timedelta'"

    conn = request.getfixturevalue(conn)
    with tm.assert_produces_warning(exp_warning, match=msg, check_stacklevel=False):
        df.to_sql(name="test_arrow", con=conn, if_exists="replace", index=False)


@pytest.mark.parametrize("conn", all_connectable)
def test_dataframe_to_sql_arrow_dtypes_missing(conn, request, nulls_fixture):
    # GH 52046
    pytest.importorskip("pyarrow")
    df = DataFrame(
        {
            "datetime": pd.array(
                [datetime(2023, 1, 1), nulls_fixture], dtype="timestamp[ns][pyarrow]"
            ),
        }
    )
    conn = request.getfixturevalue(conn)
    df.to_sql(name="test_arrow", con=conn, if_exists="replace", index=False)


@pytest.mark.parametrize("conn", all_connectable)
@pytest.mark.parametrize("method", [None, "multi"])
def test_to_sql(conn, method, test_frame1, request):
    if method == "multi" and "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="'method' not implemented for ADBC drivers", strict=True
            )
        )

    conn = request.getfixturevalue(conn)
    with pandasSQL_builder(conn, need_transaction=True) as pandasSQL:
        pandasSQL.to_sql(test_frame1, "test_frame", method=method)
        assert pandasSQL.has_table("test_frame")
    assert count_rows(conn, "test_frame") == len(test_frame1)


@pytest.mark.parametrize("conn", all_connectable)
@pytest.mark.parametrize("mode, num_row_coef", [("replace", 1), ("append", 2)])
def test_to_sql_exist(conn, mode, num_row_coef, test_frame1, request):
    conn = request.getfixturevalue(conn)
    with pandasSQL_builder(conn, need_transaction=True) as pandasSQL:
        pandasSQL.to_sql(test_frame1, "test_frame", if_exists="fail")
        pandasSQL.to_sql(test_frame1, "test_frame", if_exists=mode)
        assert pandasSQL.has_table("test_frame")
    assert count_rows(conn, "test_frame") == num_row_coef * len(test_frame1)


@pytest.mark.parametrize("conn", all_connectable)
def test_to_sql_exist_fail(conn, test_frame1, request):
    conn = request.getfixturevalue(conn)
    with pandasSQL_builder(conn, need_transaction=True) as pandasSQL:
        pandasSQL.to_sql(test_frame1, "test_frame", if_exists="fail")
        assert pandasSQL.has_table("test_frame")

        msg = "Table 'test_frame' already exists"
        with pytest.raises(ValueError, match=msg):
            pandasSQL.to_sql(test_frame1, "test_frame", if_exists="fail")


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_read_iris_query(conn, request):
    conn = request.getfixturevalue(conn)
    iris_frame = read_sql_query("SELECT * FROM iris", conn)
    check_iris_frame(iris_frame)
    iris_frame = pd.read_sql("SELECT * FROM iris", conn)
    check_iris_frame(iris_frame)
    iris_frame = pd.read_sql("SELECT * FROM iris where 0=1", conn)
    assert iris_frame.shape == (0, 5)
    assert "SepalWidth" in iris_frame.columns


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_read_iris_query_chunksize(conn, request):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="'chunksize' not implemented for ADBC drivers",
                strict=True,
            )
        )
    conn = request.getfixturevalue(conn)
    iris_frame = concat(read_sql_query("SELECT * FROM iris", conn, chunksize=7))
    check_iris_frame(iris_frame)
    iris_frame = concat(pd.read_sql("SELECT * FROM iris", conn, chunksize=7))
    check_iris_frame(iris_frame)
    iris_frame = concat(pd.read_sql("SELECT * FROM iris where 0=1", conn, chunksize=7))
    assert iris_frame.shape == (0, 5)
    assert "SepalWidth" in iris_frame.columns


@pytest.mark.parametrize("conn", sqlalchemy_connectable_iris)
def test_read_iris_query_expression_with_parameter(conn, request):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="'chunksize' not implemented for ADBC drivers",
                strict=True,
            )
        )
    conn = request.getfixturevalue(conn)
    from sqlalchemy import (
        MetaData,
        Table,
        create_engine,
        select,
    )

    metadata = MetaData()
    autoload_con = create_engine(conn) if isinstance(conn, str) else conn
    iris = Table("iris", metadata, autoload_with=autoload_con)
    iris_frame = read_sql_query(
        select(iris), conn, params={"name": "Iris-setosa", "length": 5.1}
    )
    check_iris_frame(iris_frame)
    if isinstance(conn, str):
        autoload_con.dispose()


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_read_iris_query_string_with_parameter(conn, request, sql_strings):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="'chunksize' not implemented for ADBC drivers",
                strict=True,
            )
        )

    for db, query in sql_strings["read_parameters"].items():
        if db in conn:
            break
    else:
        raise KeyError(f"No part of {conn} found in sql_strings['read_parameters']")
    conn = request.getfixturevalue(conn)
    iris_frame = read_sql_query(query, conn, params=("Iris-setosa", 5.1))
    check_iris_frame(iris_frame)


@pytest.mark.parametrize("conn", sqlalchemy_connectable_iris)
def test_read_iris_table(conn, request):
    # GH 51015 if conn = sqlite_iris_str
    conn = request.getfixturevalue(conn)
    iris_frame = read_sql_table("iris", conn)
    check_iris_frame(iris_frame)
    iris_frame = pd.read_sql("iris", conn)
    check_iris_frame(iris_frame)


@pytest.mark.parametrize("conn", sqlalchemy_connectable_iris)
def test_read_iris_table_chunksize(conn, request):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(reason="chunksize argument NotImplemented with ADBC")
        )
    conn = request.getfixturevalue(conn)
    iris_frame = concat(read_sql_table("iris", conn, chunksize=7))
    check_iris_frame(iris_frame)
    iris_frame = concat(pd.read_sql("iris", conn, chunksize=7))
    check_iris_frame(iris_frame)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_to_sql_callable(conn, test_frame1, request):
    conn = request.getfixturevalue(conn)

    check = []  # used to double check function below is really being used

    def sample(pd_table, conn, keys, data_iter):
        check.append(1)
        data = [dict(zip(keys, row)) for row in data_iter]
        conn.execute(pd_table.table.insert(), data)

    with pandasSQL_builder(conn, need_transaction=True) as pandasSQL:
        pandasSQL.to_sql(test_frame1, "test_frame", method=sample)
        assert pandasSQL.has_table("test_frame")
    assert check == [1]
    assert count_rows(conn, "test_frame") == len(test_frame1)


@pytest.mark.parametrize("conn", all_connectable_types)
def test_default_type_conversion(conn, request):
    conn_name = conn
    if conn_name == "sqlite_buildin_types":
        request.applymarker(
            pytest.mark.xfail(
                reason="sqlite_buildin connection does not implement read_sql_table"
            )
        )

    conn = request.getfixturevalue(conn)
    df = sql.read_sql_table("types", conn)

    assert issubclass(df.FloatCol.dtype.type, np.floating)
    assert issubclass(df.IntCol.dtype.type, np.integer)

    # MySQL/sqlite has no real BOOL type
    if "postgresql" in conn_name:
        assert issubclass(df.BoolCol.dtype.type, np.bool_)
    else:
        assert issubclass(df.BoolCol.dtype.type, np.integer)

    # Int column with NA values stays as float
    assert issubclass(df.IntColWithNull.dtype.type, np.floating)

    # Bool column with NA = int column with NA values => becomes float
    if "postgresql" in conn_name:
        assert issubclass(df.BoolColWithNull.dtype.type, object)
    else:
        assert issubclass(df.BoolColWithNull.dtype.type, np.floating)


@pytest.mark.parametrize("conn", mysql_connectable)
def test_read_procedure(conn, request):
    conn = request.getfixturevalue(conn)

    # GH 7324
    # Although it is more an api test, it is added to the
    # mysql tests as sqlite does not have stored procedures
    from sqlalchemy import text
    from sqlalchemy.engine import Engine

    df = DataFrame({"a": [1, 2, 3], "b": [0.1, 0.2, 0.3]})
    df.to_sql(name="test_frame", con=conn, index=False)

    proc = """DROP PROCEDURE IF EXISTS get_testdb;

    CREATE PROCEDURE get_testdb ()

    BEGIN
        SELECT * FROM test_frame;
    END"""
    proc = text(proc)
    if isinstance(conn, Engine):
        with conn.connect() as engine_conn:
            with engine_conn.begin():
                engine_conn.execute(proc)
    else:
        with conn.begin():
            conn.execute(proc)

    res1 = sql.read_sql_query("CALL get_testdb();", conn)
    tm.assert_frame_equal(df, res1)

    # test delegation to read_sql_query
    res2 = sql.read_sql("CALL get_testdb();", conn)
    tm.assert_frame_equal(df, res2)


@pytest.mark.parametrize("conn", postgresql_connectable)
@pytest.mark.parametrize("expected_count", [2, "Success!"])
def test_copy_from_callable_insertion_method(conn, expected_count, request):
    # GH 8953
    # Example in io.rst found under _io.sql.method
    # not available in sqlite, mysql
    def psql_insert_copy(table, conn, keys, data_iter):
        # gets a DBAPI connection that can provide a cursor
        dbapi_conn = conn.connection
        with dbapi_conn.cursor() as cur:
            s_buf = StringIO()
            writer = csv.writer(s_buf)
            writer.writerows(data_iter)
            s_buf.seek(0)

            columns = ", ".join([f'"{k}"' for k in keys])
            if table.schema:
                table_name = f"{table.schema}.{table.name}"
            else:
                table_name = table.name

            sql_query = f"COPY {table_name} ({columns}) FROM STDIN WITH CSV"
            cur.copy_expert(sql=sql_query, file=s_buf)
        return expected_count

    conn = request.getfixturevalue(conn)
    expected = DataFrame({"col1": [1, 2], "col2": [0.1, 0.2], "col3": ["a", "n"]})
    result_count = expected.to_sql(
        name="test_frame", con=conn, index=False, method=psql_insert_copy
    )
    # GH 46891
    if expected_count is None:
        assert result_count is None
    else:
        assert result_count == expected_count
    result = sql.read_sql_table("test_frame", conn)
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("conn", postgresql_connectable)
def test_insertion_method_on_conflict_do_nothing(conn, request):
    # GH 15988: Example in to_sql docstring
    conn = request.getfixturevalue(conn)

    from sqlalchemy.dialects.postgresql import insert
    from sqlalchemy.engine import Engine
    from sqlalchemy.sql import text

    def insert_on_conflict(table, conn, keys, data_iter):
        data = [dict(zip(keys, row)) for row in data_iter]
        stmt = (
            insert(table.table)
            .values(data)
            .on_conflict_do_nothing(index_elements=["a"])
        )
        result = conn.execute(stmt)
        return result.rowcount

    create_sql = text(
        """
    CREATE TABLE test_insert_conflict (
        a  integer PRIMARY KEY,
        b  numeric,
        c  text
    );
    """
    )
    if isinstance(conn, Engine):
        with conn.connect() as con:
            with con.begin():
                con.execute(create_sql)
    else:
        with conn.begin():
            conn.execute(create_sql)

    expected = DataFrame([[1, 2.1, "a"]], columns=list("abc"))
    expected.to_sql(
        name="test_insert_conflict", con=conn, if_exists="append", index=False
    )

    df_insert = DataFrame([[1, 3.2, "b"]], columns=list("abc"))
    inserted = df_insert.to_sql(
        name="test_insert_conflict",
        con=conn,
        index=False,
        if_exists="append",
        method=insert_on_conflict,
    )
    result = sql.read_sql_table("test_insert_conflict", conn)
    tm.assert_frame_equal(result, expected)
    assert inserted == 0

    # Cleanup
    with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
        pandasSQL.drop_table("test_insert_conflict")


@pytest.mark.parametrize("conn", all_connectable)
def test_to_sql_on_public_schema(conn, request):
    if "sqlite" in conn or "mysql" in conn:
        request.applymarker(
            pytest.mark.xfail(
                reason="test for public schema only specific to postgresql"
            )
        )

    conn = request.getfixturevalue(conn)

    test_data = DataFrame([[1, 2.1, "a"], [2, 3.1, "b"]], columns=list("abc"))
    test_data.to_sql(
        name="test_public_schema",
        con=conn,
        if_exists="append",
        index=False,
        schema="public",
    )

    df_out = sql.read_sql_table("test_public_schema", conn, schema="public")
    tm.assert_frame_equal(test_data, df_out)


@pytest.mark.parametrize("conn", mysql_connectable)
def test_insertion_method_on_conflict_update(conn, request):
    # GH 14553: Example in to_sql docstring
    conn = request.getfixturevalue(conn)

    from sqlalchemy.dialects.mysql import insert
    from sqlalchemy.engine import Engine
    from sqlalchemy.sql import text

    def insert_on_conflict(table, conn, keys, data_iter):
        data = [dict(zip(keys, row)) for row in data_iter]
        stmt = insert(table.table).values(data)
        stmt = stmt.on_duplicate_key_update(b=stmt.inserted.b, c=stmt.inserted.c)
        result = conn.execute(stmt)
        return result.rowcount

    create_sql = text(
        """
    CREATE TABLE test_insert_conflict (
        a INT PRIMARY KEY,
        b FLOAT,
        c VARCHAR(10)
    );
    """
    )
    if isinstance(conn, Engine):
        with conn.connect() as con:
            with con.begin():
                con.execute(create_sql)
    else:
        with conn.begin():
            conn.execute(create_sql)

    df = DataFrame([[1, 2.1, "a"]], columns=list("abc"))
    df.to_sql(name="test_insert_conflict", con=conn, if_exists="append", index=False)

    expected = DataFrame([[1, 3.2, "b"]], columns=list("abc"))
    inserted = expected.to_sql(
        name="test_insert_conflict",
        con=conn,
        index=False,
        if_exists="append",
        method=insert_on_conflict,
    )
    result = sql.read_sql_table("test_insert_conflict", conn)
    tm.assert_frame_equal(result, expected)
    assert inserted == 2

    # Cleanup
    with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
        pandasSQL.drop_table("test_insert_conflict")


@pytest.mark.parametrize("conn", postgresql_connectable)
def test_read_view_postgres(conn, request):
    # GH 52969
    conn = request.getfixturevalue(conn)

    from sqlalchemy.engine import Engine
    from sqlalchemy.sql import text

    table_name = f"group_{uuid.uuid4().hex}"
    view_name = f"group_view_{uuid.uuid4().hex}"

    sql_stmt = text(
        f"""
    CREATE TABLE {table_name} (
        group_id INTEGER,
        name TEXT
    );
    INSERT INTO {table_name} VALUES
        (1, 'name');
    CREATE VIEW {view_name}
    AS
    SELECT * FROM {table_name};
    """
    )
    if isinstance(conn, Engine):
        with conn.connect() as con:
            with con.begin():
                con.execute(sql_stmt)
    else:
        with conn.begin():
            conn.execute(sql_stmt)
    result = read_sql_table(view_name, conn)
    expected = DataFrame({"group_id": [1], "name": "name"})
    tm.assert_frame_equal(result, expected)


def test_read_view_sqlite(sqlite_buildin):
    # GH 52969
    create_table = """
CREATE TABLE groups (
   group_id INTEGER,
   name TEXT
);
"""
    insert_into = """
INSERT INTO groups VALUES
    (1, 'name');
"""
    create_view = """
CREATE VIEW group_view
AS
SELECT * FROM groups;
"""
    sqlite_buildin.execute(create_table)
    sqlite_buildin.execute(insert_into)
    sqlite_buildin.execute(create_view)
    result = pd.read_sql("SELECT * FROM group_view", sqlite_buildin)
    expected = DataFrame({"group_id": [1], "name": "name"})
    tm.assert_frame_equal(result, expected)


def test_execute_typeerror(sqlite_engine_iris):
    with pytest.raises(TypeError, match="pandas.io.sql.execute requires a connection"):
        with tm.assert_produces_warning(
            FutureWarning,
            match="`pandas.io.sql.execute` is deprecated and "
            "will be removed in the future version.",
        ):
            sql.execute("select * from iris", sqlite_engine_iris)


def test_execute_deprecated(sqlite_conn_iris):
    # GH50185
    with tm.assert_produces_warning(
        FutureWarning,
        match="`pandas.io.sql.execute` is deprecated and "
        "will be removed in the future version.",
    ):
        sql.execute("select * from iris", sqlite_conn_iris)


def flavor(conn_name):
    if "postgresql" in conn_name:
        return "postgresql"
    elif "sqlite" in conn_name:
        return "sqlite"
    elif "mysql" in conn_name:
        return "mysql"

    raise ValueError(f"unsupported connection: {conn_name}")


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_read_sql_iris_parameter(conn, request, sql_strings):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="'params' not implemented for ADBC drivers",
                strict=True,
            )
        )
    conn_name = conn
    conn = request.getfixturevalue(conn)
    query = sql_strings["read_parameters"][flavor(conn_name)]
    params = ("Iris-setosa", 5.1)
    with pandasSQL_builder(conn) as pandasSQL:
        with pandasSQL.run_transaction():
            iris_frame = pandasSQL.read_query(query, params=params)
    check_iris_frame(iris_frame)


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_read_sql_iris_named_parameter(conn, request, sql_strings):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="'params' not implemented for ADBC drivers",
                strict=True,
            )
        )

    conn_name = conn
    conn = request.getfixturevalue(conn)
    query = sql_strings["read_named_parameters"][flavor(conn_name)]
    params = {"name": "Iris-setosa", "length": 5.1}
    with pandasSQL_builder(conn) as pandasSQL:
        with pandasSQL.run_transaction():
            iris_frame = pandasSQL.read_query(query, params=params)
    check_iris_frame(iris_frame)


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_read_sql_iris_no_parameter_with_percent(conn, request, sql_strings):
    if "mysql" in conn or ("postgresql" in conn and "adbc" not in conn):
        request.applymarker(pytest.mark.xfail(reason="broken test"))

    conn_name = conn
    conn = request.getfixturevalue(conn)

    query = sql_strings["read_no_parameters_with_percent"][flavor(conn_name)]
    with pandasSQL_builder(conn) as pandasSQL:
        with pandasSQL.run_transaction():
            iris_frame = pandasSQL.read_query(query, params=None)
    check_iris_frame(iris_frame)


# -----------------------------------------------------------------------------
# -- Testing the public API


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_api_read_sql_view(conn, request):
    conn = request.getfixturevalue(conn)
    iris_frame = sql.read_sql_query("SELECT * FROM iris_view", conn)
    check_iris_frame(iris_frame)


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_api_read_sql_with_chunksize_no_result(conn, request):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(reason="chunksize argument NotImplemented with ADBC")
        )
    conn = request.getfixturevalue(conn)
    query = 'SELECT * FROM iris_view WHERE "SepalLength" < 0.0'
    with_batch = sql.read_sql_query(query, conn, chunksize=5)
    without_batch = sql.read_sql_query(query, conn)
    tm.assert_frame_equal(concat(with_batch), without_batch)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_to_sql(conn, request, test_frame1):
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_frame1", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_frame1")

    sql.to_sql(test_frame1, "test_frame1", conn)
    assert sql.has_table("test_frame1", conn)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_to_sql_fail(conn, request, test_frame1):
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_frame2", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_frame2")

    sql.to_sql(test_frame1, "test_frame2", conn, if_exists="fail")
    assert sql.has_table("test_frame2", conn)

    msg = "Table 'test_frame2' already exists"
    with pytest.raises(ValueError, match=msg):
        sql.to_sql(test_frame1, "test_frame2", conn, if_exists="fail")


@pytest.mark.parametrize("conn", all_connectable)
def test_api_to_sql_replace(conn, request, test_frame1):
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_frame3", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_frame3")

    sql.to_sql(test_frame1, "test_frame3", conn, if_exists="fail")
    # Add to table again
    sql.to_sql(test_frame1, "test_frame3", conn, if_exists="replace")
    assert sql.has_table("test_frame3", conn)

    num_entries = len(test_frame1)
    num_rows = count_rows(conn, "test_frame3")

    assert num_rows == num_entries


@pytest.mark.parametrize("conn", all_connectable)
def test_api_to_sql_append(conn, request, test_frame1):
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_frame4", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_frame4")

    assert sql.to_sql(test_frame1, "test_frame4", conn, if_exists="fail") == 4

    # Add to table again
    assert sql.to_sql(test_frame1, "test_frame4", conn, if_exists="append") == 4
    assert sql.has_table("test_frame4", conn)

    num_entries = 2 * len(test_frame1)
    num_rows = count_rows(conn, "test_frame4")

    assert num_rows == num_entries


@pytest.mark.parametrize("conn", all_connectable)
def test_api_to_sql_type_mapping(conn, request, test_frame3):
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_frame5", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_frame5")

    sql.to_sql(test_frame3, "test_frame5", conn, index=False)
    result = sql.read_sql("SELECT * FROM test_frame5", conn)

    tm.assert_frame_equal(test_frame3, result)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_to_sql_series(conn, request):
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_series", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_series")

    s = Series(np.arange(5, dtype="int64"), name="series")
    sql.to_sql(s, "test_series", conn, index=False)
    s2 = sql.read_sql_query("SELECT * FROM test_series", conn)
    tm.assert_frame_equal(s.to_frame(), s2)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_roundtrip(conn, request, test_frame1):
    conn_name = conn
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_frame_roundtrip", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_frame_roundtrip")

    sql.to_sql(test_frame1, "test_frame_roundtrip", con=conn)
    result = sql.read_sql_query("SELECT * FROM test_frame_roundtrip", con=conn)

    # HACK!
    if "adbc" in conn_name:
        result = result.rename(columns={"__index_level_0__": "level_0"})
    result.index = test_frame1.index
    result.set_index("level_0", inplace=True)
    result.index.astype(int)
    result.index.name = None
    tm.assert_frame_equal(result, test_frame1)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_roundtrip_chunksize(conn, request, test_frame1):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(reason="chunksize argument NotImplemented with ADBC")
        )
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_frame_roundtrip", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_frame_roundtrip")

    sql.to_sql(
        test_frame1,
        "test_frame_roundtrip",
        con=conn,
        index=False,
        chunksize=2,
    )
    result = sql.read_sql_query("SELECT * FROM test_frame_roundtrip", con=conn)
    tm.assert_frame_equal(result, test_frame1)


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_api_execute_sql(conn, request):
    # drop_sql = "DROP TABLE IF EXISTS test"  # should already be done
    conn = request.getfixturevalue(conn)
    with sql.pandasSQL_builder(conn) as pandas_sql:
        iris_results = pandas_sql.execute("SELECT * FROM iris")
        row = iris_results.fetchone()
        iris_results.close()
    assert list(row) == [5.1, 3.5, 1.4, 0.2, "Iris-setosa"]


@pytest.mark.parametrize("conn", all_connectable_types)
def test_api_date_parsing(conn, request):
    conn_name = conn
    conn = request.getfixturevalue(conn)
    # Test date parsing in read_sql
    # No Parsing
    df = sql.read_sql_query("SELECT * FROM types", conn)
    if not ("mysql" in conn_name or "postgres" in conn_name):
        assert not issubclass(df.DateCol.dtype.type, np.datetime64)

    df = sql.read_sql_query("SELECT * FROM types", conn, parse_dates=["DateCol"])
    assert issubclass(df.DateCol.dtype.type, np.datetime64)
    assert df.DateCol.tolist() == [
        Timestamp(2000, 1, 3, 0, 0, 0),
        Timestamp(2000, 1, 4, 0, 0, 0),
    ]

    df = sql.read_sql_query(
        "SELECT * FROM types",
        conn,
        parse_dates={"DateCol": "%Y-%m-%d %H:%M:%S"},
    )
    assert issubclass(df.DateCol.dtype.type, np.datetime64)
    assert df.DateCol.tolist() == [
        Timestamp(2000, 1, 3, 0, 0, 0),
        Timestamp(2000, 1, 4, 0, 0, 0),
    ]

    df = sql.read_sql_query("SELECT * FROM types", conn, parse_dates=["IntDateCol"])
    assert issubclass(df.IntDateCol.dtype.type, np.datetime64)
    assert df.IntDateCol.tolist() == [
        Timestamp(1986, 12, 25, 0, 0, 0),
        Timestamp(2013, 1, 1, 0, 0, 0),
    ]

    df = sql.read_sql_query(
        "SELECT * FROM types", conn, parse_dates={"IntDateCol": "s"}
    )
    assert issubclass(df.IntDateCol.dtype.type, np.datetime64)
    assert df.IntDateCol.tolist() == [
        Timestamp(1986, 12, 25, 0, 0, 0),
        Timestamp(2013, 1, 1, 0, 0, 0),
    ]

    df = sql.read_sql_query(
        "SELECT * FROM types",
        conn,
        parse_dates={"IntDateOnlyCol": "%Y%m%d"},
    )
    assert issubclass(df.IntDateOnlyCol.dtype.type, np.datetime64)
    assert df.IntDateOnlyCol.tolist() == [
        Timestamp("2010-10-10"),
        Timestamp("2010-12-12"),
    ]


@pytest.mark.parametrize("conn", all_connectable_types)
@pytest.mark.parametrize("error", ["ignore", "raise", "coerce"])
@pytest.mark.parametrize(
    "read_sql, text, mode",
    [
        (sql.read_sql, "SELECT * FROM types", ("sqlalchemy", "fallback")),
        (sql.read_sql, "types", ("sqlalchemy")),
        (
            sql.read_sql_query,
            "SELECT * FROM types",
            ("sqlalchemy", "fallback"),
        ),
        (sql.read_sql_table, "types", ("sqlalchemy")),
    ],
)
def test_api_custom_dateparsing_error(
    conn, request, read_sql, text, mode, error, types_data_frame
):
    conn_name = conn
    conn = request.getfixturevalue(conn)
    if text == "types" and conn_name == "sqlite_buildin_types":
        request.applymarker(
            pytest.mark.xfail(reason="failing combination of arguments")
        )

    expected = types_data_frame.astype({"DateCol": "datetime64[ns]"})

    result = read_sql(
        text,
        con=conn,
        parse_dates={
            "DateCol": {"errors": error},
        },
    )
    if "postgres" in conn_name:
        # TODO: clean up types_data_frame fixture
        result["BoolCol"] = result["BoolCol"].astype(int)
        result["BoolColWithNull"] = result["BoolColWithNull"].astype(float)

    if conn_name == "postgresql_adbc_types":
        expected = expected.astype(
            {
                "IntDateCol": "int32",
                "IntDateOnlyCol": "int32",
                "IntCol": "int32",
            }
        )

        if not pa_version_under13p0:
            # TODO: is this astype safe?
            expected["DateCol"] = expected["DateCol"].astype("datetime64[us]")

    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("conn", all_connectable_types)
def test_api_date_and_index(conn, request):
    # Test case where same column appears in parse_date and index_col
    conn = request.getfixturevalue(conn)
    df = sql.read_sql_query(
        "SELECT * FROM types",
        conn,
        index_col="DateCol",
        parse_dates=["DateCol", "IntDateCol"],
    )

    assert issubclass(df.index.dtype.type, np.datetime64)
    assert issubclass(df.IntDateCol.dtype.type, np.datetime64)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_timedelta(conn, request):
    # see #6921
    conn_name = conn
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_timedelta", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_timedelta")

    df = to_timedelta(Series(["00:00:01", "00:00:03"], name="foo")).to_frame()

    if conn_name == "sqlite_adbc_conn":
        request.node.add_marker(
            pytest.mark.xfail(
                reason="sqlite ADBC driver doesn't implement timedelta",
            )
        )

    if "adbc" in conn_name:
        if pa_version_under14p1:
            exp_warning = DeprecationWarning
        else:
            exp_warning = None
    else:
        exp_warning = UserWarning

    with tm.assert_produces_warning(exp_warning, check_stacklevel=False):
        result_count = df.to_sql(name="test_timedelta", con=conn)
    assert result_count == 2
    result = sql.read_sql_query("SELECT * FROM test_timedelta", conn)

    if conn_name == "postgresql_adbc_conn":
        # TODO: Postgres stores an INTERVAL, which ADBC reads as a Month-Day-Nano
        # Interval; the default pandas type mapper maps this to a DateOffset
        # but maybe we should try and restore the timedelta here?
        expected = Series(
            [
                pd.DateOffset(months=0, days=0, microseconds=1000000, nanoseconds=0),
                pd.DateOffset(months=0, days=0, microseconds=3000000, nanoseconds=0),
            ],
            name="foo",
        )
    else:
        expected = df["foo"].astype("int64")
    tm.assert_series_equal(result["foo"], expected)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_complex_raises(conn, request):
    conn_name = conn
    conn = request.getfixturevalue(conn)
    df = DataFrame({"a": [1 + 1j, 2j]})

    if "adbc" in conn_name:
        msg = "datatypes not supported"
    else:
        msg = "Complex datatypes not supported"
    with pytest.raises(ValueError, match=msg):
        assert df.to_sql("test_complex", con=conn) is None


@pytest.mark.parametrize("conn", all_connectable)
@pytest.mark.parametrize(
    "index_name,index_label,expected",
    [
        # no index name, defaults to 'index'
        (None, None, "index"),
        # specifying index_label
        (None, "other_label", "other_label"),
        # using the index name
        ("index_name", None, "index_name"),
        # has index name, but specifying index_label
        ("index_name", "other_label", "other_label"),
        # index name is integer
        (0, None, "0"),
        # index name is None but index label is integer
        (None, 0, "0"),
    ],
)
def test_api_to_sql_index_label(conn, request, index_name, index_label, expected):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(reason="index_label argument NotImplemented with ADBC")
        )
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_index_label", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_index_label")

    temp_frame = DataFrame({"col1": range(4)})
    temp_frame.index.name = index_name
    query = "SELECT * FROM test_index_label"
    sql.to_sql(temp_frame, "test_index_label", conn, index_label=index_label)
    frame = sql.read_sql_query(query, conn)
    assert frame.columns[0] == expected


@pytest.mark.parametrize("conn", all_connectable)
def test_api_to_sql_index_label_multiindex(conn, request):
    conn_name = conn
    if "mysql" in conn_name:
        request.applymarker(
            pytest.mark.xfail(
                reason="MySQL can fail using TEXT without length as key", strict=False
            )
        )
    elif "adbc" in conn_name:
        request.node.add_marker(
            pytest.mark.xfail(reason="index_label argument NotImplemented with ADBC")
        )

    conn = request.getfixturevalue(conn)
    if sql.has_table("test_index_label", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_index_label")

    expected_row_count = 4
    temp_frame = DataFrame(
        {"col1": range(4)},
        index=MultiIndex.from_product([("A0", "A1"), ("B0", "B1")]),
    )

    # no index name, defaults to 'level_0' and 'level_1'
    result = sql.to_sql(temp_frame, "test_index_label", conn)
    assert result == expected_row_count
    frame = sql.read_sql_query("SELECT * FROM test_index_label", conn)
    assert frame.columns[0] == "level_0"
    assert frame.columns[1] == "level_1"

    # specifying index_label
    result = sql.to_sql(
        temp_frame,
        "test_index_label",
        conn,
        if_exists="replace",
        index_label=["A", "B"],
    )
    assert result == expected_row_count
    frame = sql.read_sql_query("SELECT * FROM test_index_label", conn)
    assert frame.columns[:2].tolist() == ["A", "B"]

    # using the index name
    temp_frame.index.names = ["A", "B"]
    result = sql.to_sql(temp_frame, "test_index_label", conn, if_exists="replace")
    assert result == expected_row_count
    frame = sql.read_sql_query("SELECT * FROM test_index_label", conn)
    assert frame.columns[:2].tolist() == ["A", "B"]

    # has index name, but specifying index_label
    result = sql.to_sql(
        temp_frame,
        "test_index_label",
        conn,
        if_exists="replace",
        index_label=["C", "D"],
    )
    assert result == expected_row_count
    frame = sql.read_sql_query("SELECT * FROM test_index_label", conn)
    assert frame.columns[:2].tolist() == ["C", "D"]

    msg = "Length of 'index_label' should match number of levels, which is 2"
    with pytest.raises(ValueError, match=msg):
        sql.to_sql(
            temp_frame,
            "test_index_label",
            conn,
            if_exists="replace",
            index_label="C",
        )


@pytest.mark.parametrize("conn", all_connectable)
def test_api_multiindex_roundtrip(conn, request):
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_multiindex_roundtrip", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_multiindex_roundtrip")

    df = DataFrame.from_records(
        [(1, 2.1, "line1"), (2, 1.5, "line2")],
        columns=["A", "B", "C"],
        index=["A", "B"],
    )

    df.to_sql(name="test_multiindex_roundtrip", con=conn)
    result = sql.read_sql_query(
        "SELECT * FROM test_multiindex_roundtrip", conn, index_col=["A", "B"]
    )
    tm.assert_frame_equal(df, result, check_index_type=True)


@pytest.mark.parametrize("conn", all_connectable)
@pytest.mark.parametrize(
    "dtype",
    [
        None,
        int,
        float,
        {"A": int, "B": float},
    ],
)
def test_api_dtype_argument(conn, request, dtype):
    # GH10285 Add dtype argument to read_sql_query
    conn_name = conn
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_dtype_argument", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_dtype_argument")

    df = DataFrame([[1.2, 3.4], [5.6, 7.8]], columns=["A", "B"])
    assert df.to_sql(name="test_dtype_argument", con=conn) == 2

    expected = df.astype(dtype)

    if "postgres" in conn_name:
        query = 'SELECT "A", "B" FROM test_dtype_argument'
    else:
        query = "SELECT A, B FROM test_dtype_argument"
    result = sql.read_sql_query(query, con=conn, dtype=dtype)

    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_integer_col_names(conn, request):
    conn = request.getfixturevalue(conn)
    df = DataFrame([[1, 2], [3, 4]], columns=[0, 1])
    sql.to_sql(df, "test_frame_integer_col_names", conn, if_exists="replace")


@pytest.mark.parametrize("conn", all_connectable)
def test_api_get_schema(conn, request, test_frame1):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="'get_schema' not implemented for ADBC drivers",
                strict=True,
            )
        )
    conn = request.getfixturevalue(conn)
    create_sql = sql.get_schema(test_frame1, "test", con=conn)
    assert "CREATE" in create_sql


@pytest.mark.parametrize("conn", all_connectable)
def test_api_get_schema_with_schema(conn, request, test_frame1):
    # GH28486
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="'get_schema' not implemented for ADBC drivers",
                strict=True,
            )
        )
    conn = request.getfixturevalue(conn)
    create_sql = sql.get_schema(test_frame1, "test", con=conn, schema="pypi")
    assert "CREATE TABLE pypi." in create_sql


@pytest.mark.parametrize("conn", all_connectable)
def test_api_get_schema_dtypes(conn, request):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="'get_schema' not implemented for ADBC drivers",
                strict=True,
            )
        )
    conn_name = conn
    conn = request.getfixturevalue(conn)
    float_frame = DataFrame({"a": [1.1, 1.2], "b": [2.1, 2.2]})

    if conn_name == "sqlite_buildin":
        dtype = "INTEGER"
    else:
        from sqlalchemy import Integer

        dtype = Integer
    create_sql = sql.get_schema(float_frame, "test", con=conn, dtype={"b": dtype})
    assert "CREATE" in create_sql
    assert "INTEGER" in create_sql


@pytest.mark.parametrize("conn", all_connectable)
def test_api_get_schema_keys(conn, request, test_frame1):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="'get_schema' not implemented for ADBC drivers",
                strict=True,
            )
        )
    conn_name = conn
    conn = request.getfixturevalue(conn)
    frame = DataFrame({"Col1": [1.1, 1.2], "Col2": [2.1, 2.2]})
    create_sql = sql.get_schema(frame, "test", con=conn, keys="Col1")

    if "mysql" in conn_name:
        constraint_sentence = "CONSTRAINT test_pk PRIMARY KEY (`Col1`)"
    else:
        constraint_sentence = 'CONSTRAINT test_pk PRIMARY KEY ("Col1")'
    assert constraint_sentence in create_sql

    # multiple columns as key (GH10385)
    create_sql = sql.get_schema(test_frame1, "test", con=conn, keys=["A", "B"])
    if "mysql" in conn_name:
        constraint_sentence = "CONSTRAINT test_pk PRIMARY KEY (`A`, `B`)"
    else:
        constraint_sentence = 'CONSTRAINT test_pk PRIMARY KEY ("A", "B")'
    assert constraint_sentence in create_sql


@pytest.mark.parametrize("conn", all_connectable)
def test_api_chunksize_read(conn, request):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(reason="chunksize argument NotImplemented with ADBC")
        )
    conn_name = conn
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_chunksize", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_chunksize")

    df = DataFrame(
        np.random.default_rng(2).standard_normal((22, 5)), columns=list("abcde")
    )
    df.to_sql(name="test_chunksize", con=conn, index=False)

    # reading the query in one time
    res1 = sql.read_sql_query("select * from test_chunksize", conn)

    # reading the query in chunks with read_sql_query
    res2 = DataFrame()
    i = 0
    sizes = [5, 5, 5, 5, 2]

    for chunk in sql.read_sql_query("select * from test_chunksize", conn, chunksize=5):
        res2 = concat([res2, chunk], ignore_index=True)
        assert len(chunk) == sizes[i]
        i += 1

    tm.assert_frame_equal(res1, res2)

    # reading the query in chunks with read_sql_query
    if conn_name == "sqlite_buildin":
        with pytest.raises(NotImplementedError, match=""):
            sql.read_sql_table("test_chunksize", conn, chunksize=5)
    else:
        res3 = DataFrame()
        i = 0
        sizes = [5, 5, 5, 5, 2]

        for chunk in sql.read_sql_table("test_chunksize", conn, chunksize=5):
            res3 = concat([res3, chunk], ignore_index=True)
            assert len(chunk) == sizes[i]
            i += 1

        tm.assert_frame_equal(res1, res3)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_categorical(conn, request):
    if conn == "postgresql_adbc_conn":
        adbc = import_optional_dependency("adbc_driver_postgresql", errors="ignore")
        if adbc is not None and Version(adbc.__version__) < Version("0.9.0"):
            request.node.add_marker(
                pytest.mark.xfail(
                    reason="categorical dtype not implemented for ADBC postgres driver",
                    strict=True,
                )
            )
    # GH8624
    # test that categorical gets written correctly as dense column
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_categorical", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_categorical")

    df = DataFrame(
        {
            "person_id": [1, 2, 3],
            "person_name": ["John P. Doe", "Jane Dove", "John P. Doe"],
        }
    )
    df2 = df.copy()
    df2["person_name"] = df2["person_name"].astype("category")

    df2.to_sql(name="test_categorical", con=conn, index=False)
    res = sql.read_sql_query("SELECT * FROM test_categorical", conn)

    tm.assert_frame_equal(res, df)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_unicode_column_name(conn, request):
    # GH 11431
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_unicode", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_unicode")

    df = DataFrame([[1, 2], [3, 4]], columns=["\xe9", "b"])
    df.to_sql(name="test_unicode", con=conn, index=False)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_escaped_table_name(conn, request):
    # GH 13206
    conn_name = conn
    conn = request.getfixturevalue(conn)
    if sql.has_table("d1187b08-4943-4c8d-a7f6", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("d1187b08-4943-4c8d-a7f6")

    df = DataFrame({"A": [0, 1, 2], "B": [0.2, np.nan, 5.6]})
    df.to_sql(name="d1187b08-4943-4c8d-a7f6", con=conn, index=False)

    if "postgres" in conn_name:
        query = 'SELECT * FROM "d1187b08-4943-4c8d-a7f6"'
    else:
        query = "SELECT * FROM `d1187b08-4943-4c8d-a7f6`"
    res = sql.read_sql_query(query, conn)

    tm.assert_frame_equal(res, df)


@pytest.mark.parametrize("conn", all_connectable)
def test_api_read_sql_duplicate_columns(conn, request):
    # GH#53117
    if "adbc" in conn:
        pa = pytest.importorskip("pyarrow")
        if not (
            Version(pa.__version__) >= Version("16.0")
            and conn in ["sqlite_adbc_conn", "postgresql_adbc_conn"]
        ):
            request.node.add_marker(
                pytest.mark.xfail(
                    reason="pyarrow->pandas throws ValueError", strict=True
                )
            )
    conn = request.getfixturevalue(conn)
    if sql.has_table("test_table", conn):
        with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
            pandasSQL.drop_table("test_table")

    df = DataFrame({"a": [1, 2, 3], "b": [0.1, 0.2, 0.3], "c": 1})
    df.to_sql(name="test_table", con=conn, index=False)

    result = pd.read_sql("SELECT a, b, a +1 as a, c FROM test_table", conn)
    expected = DataFrame(
        [[1, 0.1, 2, 1], [2, 0.2, 3, 1], [3, 0.3, 4, 1]],
        columns=["a", "b", "a", "c"],
    )
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("conn", all_connectable)
def test_read_table_columns(conn, request, test_frame1):
    # test columns argument in read_table
    conn_name = conn
    if conn_name == "sqlite_buildin":
        request.applymarker(pytest.mark.xfail(reason="Not Implemented"))

    conn = request.getfixturevalue(conn)
    sql.to_sql(test_frame1, "test_frame", conn)

    cols = ["A", "B"]

    result = sql.read_sql_table("test_frame", conn, columns=cols)
    assert result.columns.tolist() == cols


@pytest.mark.parametrize("conn", all_connectable)
def test_read_table_index_col(conn, request, test_frame1):
    # test columns argument in read_table
    conn_name = conn
    if conn_name == "sqlite_buildin":
        request.applymarker(pytest.mark.xfail(reason="Not Implemented"))

    conn = request.getfixturevalue(conn)
    sql.to_sql(test_frame1, "test_frame", conn)

    result = sql.read_sql_table("test_frame", conn, index_col="index")
    assert result.index.names == ["index"]

    result = sql.read_sql_table("test_frame", conn, index_col=["A", "B"])
    assert result.index.names == ["A", "B"]

    result = sql.read_sql_table(
        "test_frame", conn, index_col=["A", "B"], columns=["C", "D"]
    )
    assert result.index.names == ["A", "B"]
    assert result.columns.tolist() == ["C", "D"]


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_read_sql_delegate(conn, request):
    if conn == "sqlite_buildin_iris":
        request.applymarker(
            pytest.mark.xfail(
                reason="sqlite_buildin connection does not implement read_sql_table"
            )
        )

    conn = request.getfixturevalue(conn)
    iris_frame1 = sql.read_sql_query("SELECT * FROM iris", conn)
    iris_frame2 = sql.read_sql("SELECT * FROM iris", conn)
    tm.assert_frame_equal(iris_frame1, iris_frame2)

    iris_frame1 = sql.read_sql_table("iris", conn)
    iris_frame2 = sql.read_sql("iris", conn)
    tm.assert_frame_equal(iris_frame1, iris_frame2)


def test_not_reflect_all_tables(sqlite_conn):
    conn = sqlite_conn
    from sqlalchemy import text
    from sqlalchemy.engine import Engine

    # create invalid table
    query_list = [
        text("CREATE TABLE invalid (x INTEGER, y UNKNOWN);"),
        text("CREATE TABLE other_table (x INTEGER, y INTEGER);"),
    ]

    for query in query_list:
        if isinstance(conn, Engine):
            with conn.connect() as conn:
                with conn.begin():
                    conn.execute(query)
        else:
            with conn.begin():
                conn.execute(query)

    with tm.assert_produces_warning(None):
        sql.read_sql_table("other_table", conn)
        sql.read_sql_query("SELECT * FROM other_table", conn)


@pytest.mark.parametrize("conn", all_connectable)
def test_warning_case_insensitive_table_name(conn, request, test_frame1):
    conn_name = conn
    if conn_name == "sqlite_buildin" or "adbc" in conn_name:
        request.applymarker(pytest.mark.xfail(reason="Does not raise warning"))

    conn = request.getfixturevalue(conn)
    # see gh-7815
    with tm.assert_produces_warning(
        UserWarning,
        match=(
            r"The provided table name 'TABLE1' is not found exactly as such in "
            r"the database after writing the table, possibly due to case "
            r"sensitivity issues. Consider using lower case table names."
        ),
    ):
        with sql.SQLDatabase(conn) as db:
            db.check_case_sensitive("TABLE1", "")

    # Test that the warning is certainly NOT triggered in a normal case.
    with tm.assert_produces_warning(None):
        test_frame1.to_sql(name="CaseSensitive", con=conn)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_sqlalchemy_type_mapping(conn, request):
    conn = request.getfixturevalue(conn)
    from sqlalchemy import TIMESTAMP

    # Test Timestamp objects (no datetime64 because of timezone) (GH9085)
    df = DataFrame(
        {"time": to_datetime(["2014-12-12 01:54", "2014-12-11 02:54"], utc=True)}
    )
    with sql.SQLDatabase(conn) as db:
        table = sql.SQLTable("test_type", db, frame=df)
        # GH 9086: TIMESTAMP is the suggested type for datetimes with timezones
        assert isinstance(table.table.c["time"].type, TIMESTAMP)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
@pytest.mark.parametrize(
    "integer, expected",
    [
        ("int8", "SMALLINT"),
        ("Int8", "SMALLINT"),
        ("uint8", "SMALLINT"),
        ("UInt8", "SMALLINT"),
        ("int16", "SMALLINT"),
        ("Int16", "SMALLINT"),
        ("uint16", "INTEGER"),
        ("UInt16", "INTEGER"),
        ("int32", "INTEGER"),
        ("Int32", "INTEGER"),
        ("uint32", "BIGINT"),
        ("UInt32", "BIGINT"),
        ("int64", "BIGINT"),
        ("Int64", "BIGINT"),
        (int, "BIGINT" if np.dtype(int).name == "int64" else "INTEGER"),
    ],
)
def test_sqlalchemy_integer_mapping(conn, request, integer, expected):
    # GH35076 Map pandas integer to optimal SQLAlchemy integer type
    conn = request.getfixturevalue(conn)
    df = DataFrame([0, 1], columns=["a"], dtype=integer)
    with sql.SQLDatabase(conn) as db:
        table = sql.SQLTable("test_type", db, frame=df)

        result = str(table.table.c.a.type)
    assert result == expected


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
@pytest.mark.parametrize("integer", ["uint64", "UInt64"])
def test_sqlalchemy_integer_overload_mapping(conn, request, integer):
    conn = request.getfixturevalue(conn)
    # GH35076 Map pandas integer to optimal SQLAlchemy integer type
    df = DataFrame([0, 1], columns=["a"], dtype=integer)
    with sql.SQLDatabase(conn) as db:
        with pytest.raises(
            ValueError, match="Unsigned 64 bit integer datatype is not supported"
        ):
            sql.SQLTable("test_type", db, frame=df)


@pytest.mark.parametrize("conn", all_connectable)
def test_database_uri_string(conn, request, test_frame1):
    pytest.importorskip("sqlalchemy")
    conn = request.getfixturevalue(conn)
    # Test read_sql and .to_sql method with a database URI (GH10654)
    # db_uri = 'sqlite:///:memory:' # raises
    # sqlalchemy.exc.OperationalError: (sqlite3.OperationalError) near
    # "iris": syntax error [SQL: 'iris']
    with tm.ensure_clean() as name:
        db_uri = "sqlite:///" + name
        table = "iris"
        test_frame1.to_sql(name=table, con=db_uri, if_exists="replace", index=False)
        test_frame2 = sql.read_sql(table, db_uri)
        test_frame3 = sql.read_sql_table(table, db_uri)
        query = "SELECT * FROM iris"
        test_frame4 = sql.read_sql_query(query, db_uri)
    tm.assert_frame_equal(test_frame1, test_frame2)
    tm.assert_frame_equal(test_frame1, test_frame3)
    tm.assert_frame_equal(test_frame1, test_frame4)


@td.skip_if_installed("pg8000")
@pytest.mark.parametrize("conn", all_connectable)
def test_pg8000_sqlalchemy_passthrough_error(conn, request):
    pytest.importorskip("sqlalchemy")
    conn = request.getfixturevalue(conn)
    # using driver that will not be installed on CI to trigger error
    # in sqlalchemy.create_engine -> test passing of this error to user
    db_uri = "postgresql+pg8000://user:pass@host/dbname"
    with pytest.raises(ImportError, match="pg8000"):
        sql.read_sql("select * from table", db_uri)


@pytest.mark.parametrize("conn", sqlalchemy_connectable_iris)
def test_query_by_text_obj(conn, request):
    # WIP : GH10846
    conn_name = conn
    conn = request.getfixturevalue(conn)
    from sqlalchemy import text

    if "postgres" in conn_name:
        name_text = text('select * from iris where "Name"=:name')
    else:
        name_text = text("select * from iris where name=:name")
    iris_df = sql.read_sql(name_text, conn, params={"name": "Iris-versicolor"})
    all_names = set(iris_df["Name"])
    assert all_names == {"Iris-versicolor"}


@pytest.mark.parametrize("conn", sqlalchemy_connectable_iris)
def test_query_by_select_obj(conn, request):
    conn = request.getfixturevalue(conn)
    # WIP : GH10846
    from sqlalchemy import (
        bindparam,
        select,
    )

    iris = iris_table_metadata()
    name_select = select(iris).where(iris.c.Name == bindparam("name"))
    iris_df = sql.read_sql(name_select, conn, params={"name": "Iris-setosa"})
    all_names = set(iris_df["Name"])
    assert all_names == {"Iris-setosa"}


@pytest.mark.parametrize("conn", all_connectable)
def test_column_with_percentage(conn, request):
    # GH 37157
    conn_name = conn
    if conn_name == "sqlite_buildin":
        request.applymarker(pytest.mark.xfail(reason="Not Implemented"))

    conn = request.getfixturevalue(conn)
    df = DataFrame({"A": [0, 1, 2], "%_variation": [3, 4, 5]})
    df.to_sql(name="test_column_percentage", con=conn, index=False)

    res = sql.read_sql_table("test_column_percentage", conn)

    tm.assert_frame_equal(res, df)


def test_sql_open_close(test_frame3):
    # Test if the IO in the database still work if the connection closed
    # between the writing and reading (as in many real situations).

    with tm.ensure_clean() as name:
        with closing(sqlite3.connect(name)) as conn:
            assert sql.to_sql(test_frame3, "test_frame3_legacy", conn, index=False) == 4

        with closing(sqlite3.connect(name)) as conn:
            result = sql.read_sql_query("SELECT * FROM test_frame3_legacy;", conn)

    tm.assert_frame_equal(test_frame3, result)


@td.skip_if_installed("sqlalchemy")
def test_con_string_import_error():
    conn = "mysql://root@localhost/pandas"
    msg = "Using URI string without sqlalchemy installed"
    with pytest.raises(ImportError, match=msg):
        sql.read_sql("SELECT * FROM iris", conn)


@td.skip_if_installed("sqlalchemy")
def test_con_unknown_dbapi2_class_does_not_error_without_sql_alchemy_installed():
    class MockSqliteConnection:
        def __init__(self, *args, **kwargs) -> None:
            self.conn = sqlite3.Connection(*args, **kwargs)

        def __getattr__(self, name):
            return getattr(self.conn, name)

        def close(self):
            self.conn.close()

    with contextlib.closing(MockSqliteConnection(":memory:")) as conn:
        with tm.assert_produces_warning(UserWarning):
            sql.read_sql("SELECT 1", conn)


def test_sqlite_read_sql_delegate(sqlite_buildin_iris):
    conn = sqlite_buildin_iris
    iris_frame1 = sql.read_sql_query("SELECT * FROM iris", conn)
    iris_frame2 = sql.read_sql("SELECT * FROM iris", conn)
    tm.assert_frame_equal(iris_frame1, iris_frame2)

    msg = "Execution failed on sql 'iris': near \"iris\": syntax error"
    with pytest.raises(sql.DatabaseError, match=msg):
        sql.read_sql("iris", conn)


def test_get_schema2(test_frame1):
    # without providing a connection object (available for backwards comp)
    create_sql = sql.get_schema(test_frame1, "test")
    assert "CREATE" in create_sql


def test_sqlite_type_mapping(sqlite_buildin):
    # Test Timestamp objects (no datetime64 because of timezone) (GH9085)
    conn = sqlite_buildin
    df = DataFrame(
        {"time": to_datetime(["2014-12-12 01:54", "2014-12-11 02:54"], utc=True)}
    )
    db = sql.SQLiteDatabase(conn)
    table = sql.SQLiteTable("test_type", db, frame=df)
    schema = table.sql_schema()
    for col in schema.split("\n"):
        if col.split()[0].strip('"') == "time":
            assert col.split()[1] == "TIMESTAMP"


# -----------------------------------------------------------------------------
# -- Database flavor specific tests


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_create_table(conn, request):
    if conn == "sqlite_str":
        pytest.skip("sqlite_str has no inspection system")

    conn = request.getfixturevalue(conn)

    from sqlalchemy import inspect

    temp_frame = DataFrame({"one": [1.0, 2.0, 3.0, 4.0], "two": [4.0, 3.0, 2.0, 1.0]})
    with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
        assert pandasSQL.to_sql(temp_frame, "temp_frame") == 4

    insp = inspect(conn)
    assert insp.has_table("temp_frame")

    # Cleanup
    with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
        pandasSQL.drop_table("temp_frame")


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_drop_table(conn, request):
    if conn == "sqlite_str":
        pytest.skip("sqlite_str has no inspection system")

    conn = request.getfixturevalue(conn)

    from sqlalchemy import inspect

    temp_frame = DataFrame({"one": [1.0, 2.0, 3.0, 4.0], "two": [4.0, 3.0, 2.0, 1.0]})
    with sql.SQLDatabase(conn) as pandasSQL:
        with pandasSQL.run_transaction():
            assert pandasSQL.to_sql(temp_frame, "temp_frame") == 4

        insp = inspect(conn)
        assert insp.has_table("temp_frame")

        with pandasSQL.run_transaction():
            pandasSQL.drop_table("temp_frame")
        try:
            insp.clear_cache()  # needed with SQLAlchemy 2.0, unavailable prior
        except AttributeError:
            pass
        assert not insp.has_table("temp_frame")


@pytest.mark.parametrize("conn", all_connectable)
def test_roundtrip(conn, request, test_frame1):
    if conn == "sqlite_str":
        pytest.skip("sqlite_str has no inspection system")

    conn_name = conn
    conn = request.getfixturevalue(conn)
    pandasSQL = pandasSQL_builder(conn)
    with pandasSQL.run_transaction():
        assert pandasSQL.to_sql(test_frame1, "test_frame_roundtrip") == 4
        result = pandasSQL.read_query("SELECT * FROM test_frame_roundtrip")

    if "adbc" in conn_name:
        result = result.rename(columns={"__index_level_0__": "level_0"})
    result.set_index("level_0", inplace=True)
    # result.index.astype(int)

    result.index.name = None

    tm.assert_frame_equal(result, test_frame1)


@pytest.mark.parametrize("conn", all_connectable_iris)
def test_execute_sql(conn, request):
    conn = request.getfixturevalue(conn)
    with pandasSQL_builder(conn) as pandasSQL:
        with pandasSQL.run_transaction():
            iris_results = pandasSQL.execute("SELECT * FROM iris")
            row = iris_results.fetchone()
            iris_results.close()
    assert list(row) == [5.1, 3.5, 1.4, 0.2, "Iris-setosa"]


@pytest.mark.parametrize("conn", sqlalchemy_connectable_iris)
def test_sqlalchemy_read_table(conn, request):
    conn = request.getfixturevalue(conn)
    iris_frame = sql.read_sql_table("iris", con=conn)
    check_iris_frame(iris_frame)


@pytest.mark.parametrize("conn", sqlalchemy_connectable_iris)
def test_sqlalchemy_read_table_columns(conn, request):
    conn = request.getfixturevalue(conn)
    iris_frame = sql.read_sql_table(
        "iris", con=conn, columns=["SepalLength", "SepalLength"]
    )
    tm.assert_index_equal(iris_frame.columns, Index(["SepalLength", "SepalLength__1"]))


@pytest.mark.parametrize("conn", sqlalchemy_connectable_iris)
def test_read_table_absent_raises(conn, request):
    conn = request.getfixturevalue(conn)
    msg = "Table this_doesnt_exist not found"
    with pytest.raises(ValueError, match=msg):
        sql.read_sql_table("this_doesnt_exist", con=conn)


@pytest.mark.parametrize("conn", sqlalchemy_connectable_types)
def test_sqlalchemy_default_type_conversion(conn, request):
    conn_name = conn
    if conn_name == "sqlite_str":
        pytest.skip("types tables not created in sqlite_str fixture")
    elif "mysql" in conn_name or "sqlite" in conn_name:
        request.applymarker(
            pytest.mark.xfail(reason="boolean dtype not inferred properly")
        )

    conn = request.getfixturevalue(conn)
    df = sql.read_sql_table("types", conn)

    assert issubclass(df.FloatCol.dtype.type, np.floating)
    assert issubclass(df.IntCol.dtype.type, np.integer)
    assert issubclass(df.BoolCol.dtype.type, np.bool_)

    # Int column with NA values stays as float
    assert issubclass(df.IntColWithNull.dtype.type, np.floating)
    # Bool column with NA values becomes object
    assert issubclass(df.BoolColWithNull.dtype.type, object)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_bigint(conn, request):
    # int64 should be converted to BigInteger, GH7433
    conn = request.getfixturevalue(conn)
    df = DataFrame(data={"i64": [2**62]})
    assert df.to_sql(name="test_bigint", con=conn, index=False) == 1
    result = sql.read_sql_table("test_bigint", conn)

    tm.assert_frame_equal(df, result)


@pytest.mark.parametrize("conn", sqlalchemy_connectable_types)
def test_default_date_load(conn, request):
    conn_name = conn
    if conn_name == "sqlite_str":
        pytest.skip("types tables not created in sqlite_str fixture")
    elif "sqlite" in conn_name:
        request.applymarker(
            pytest.mark.xfail(reason="sqlite does not read date properly")
        )

    conn = request.getfixturevalue(conn)
    df = sql.read_sql_table("types", conn)

    assert issubclass(df.DateCol.dtype.type, np.datetime64)


@pytest.mark.parametrize("conn", postgresql_connectable)
@pytest.mark.parametrize("parse_dates", [None, ["DateColWithTz"]])
def test_datetime_with_timezone_query(conn, request, parse_dates):
    # edge case that converts postgresql datetime with time zone types
    # to datetime64[ns,psycopg2.tz.FixedOffsetTimezone..], which is ok
    # but should be more natural, so coerce to datetime64[ns] for now
    conn = request.getfixturevalue(conn)
    expected = create_and_load_postgres_datetz(conn)

    # GH11216
    df = read_sql_query("select * from datetz", conn, parse_dates=parse_dates)
    col = df.DateColWithTz
    tm.assert_series_equal(col, expected)


@pytest.mark.parametrize("conn", postgresql_connectable)
def test_datetime_with_timezone_query_chunksize(conn, request):
    conn = request.getfixturevalue(conn)
    expected = create_and_load_postgres_datetz(conn)

    df = concat(
        list(read_sql_query("select * from datetz", conn, chunksize=1)),
        ignore_index=True,
    )
    col = df.DateColWithTz
    tm.assert_series_equal(col, expected)


@pytest.mark.parametrize("conn", postgresql_connectable)
def test_datetime_with_timezone_table(conn, request):
    conn = request.getfixturevalue(conn)
    expected = create_and_load_postgres_datetz(conn)
    result = sql.read_sql_table("datetz", conn)
    tm.assert_frame_equal(result, expected.to_frame())


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_datetime_with_timezone_roundtrip(conn, request):
    conn_name = conn
    conn = request.getfixturevalue(conn)
    # GH 9086
    # Write datetimetz data to a db and read it back
    # For dbs that support timestamps with timezones, should get back UTC
    # otherwise naive data should be returned
    expected = DataFrame(
        {"A": date_range("2013-01-01 09:00:00", periods=3, tz="US/Pacific")}
    )
    assert expected.to_sql(name="test_datetime_tz", con=conn, index=False) == 3

    if "postgresql" in conn_name:
        # SQLAlchemy "timezones" (i.e. offsets) are coerced to UTC
        expected["A"] = expected["A"].dt.tz_convert("UTC")
    else:
        # Otherwise, timestamps are returned as local, naive
        expected["A"] = expected["A"].dt.tz_localize(None)

    result = sql.read_sql_table("test_datetime_tz", conn)
    tm.assert_frame_equal(result, expected)

    result = sql.read_sql_query("SELECT * FROM test_datetime_tz", conn)
    if "sqlite" in conn_name:
        # read_sql_query does not return datetime type like read_sql_table
        assert isinstance(result.loc[0, "A"], str)
        result["A"] = to_datetime(result["A"])
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_out_of_bounds_datetime(conn, request):
    # GH 26761
    conn = request.getfixturevalue(conn)
    data = DataFrame({"date": datetime(9999, 1, 1)}, index=[0])
    assert data.to_sql(name="test_datetime_obb", con=conn, index=False) == 1
    result = sql.read_sql_table("test_datetime_obb", conn)
    expected = DataFrame([pd.NaT], columns=["date"])
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_naive_datetimeindex_roundtrip(conn, request):
    # GH 23510
    # Ensure that a naive DatetimeIndex isn't converted to UTC
    conn = request.getfixturevalue(conn)
    dates = date_range("2018-01-01", periods=5, freq="6h")._with_freq(None)
    expected = DataFrame({"nums": range(5)}, index=dates)
    assert expected.to_sql(name="foo_table", con=conn, index_label="info_date") == 5
    result = sql.read_sql_table("foo_table", conn, index_col="info_date")
    # result index with gain a name from a set_index operation; expected
    tm.assert_frame_equal(result, expected, check_names=False)


@pytest.mark.parametrize("conn", sqlalchemy_connectable_types)
def test_date_parsing(conn, request):
    # No Parsing
    conn_name = conn
    conn = request.getfixturevalue(conn)
    df = sql.read_sql_table("types", conn)
    expected_type = object if "sqlite" in conn_name else np.datetime64
    assert issubclass(df.DateCol.dtype.type, expected_type)

    df = sql.read_sql_table("types", conn, parse_dates=["DateCol"])
    assert issubclass(df.DateCol.dtype.type, np.datetime64)

    df = sql.read_sql_table("types", conn, parse_dates={"DateCol": "%Y-%m-%d %H:%M:%S"})
    assert issubclass(df.DateCol.dtype.type, np.datetime64)

    df = sql.read_sql_table(
        "types",
        conn,
        parse_dates={"DateCol": {"format": "%Y-%m-%d %H:%M:%S"}},
    )
    assert issubclass(df.DateCol.dtype.type, np.datetime64)

    df = sql.read_sql_table("types", conn, parse_dates=["IntDateCol"])
    assert issubclass(df.IntDateCol.dtype.type, np.datetime64)

    df = sql.read_sql_table("types", conn, parse_dates={"IntDateCol": "s"})
    assert issubclass(df.IntDateCol.dtype.type, np.datetime64)

    df = sql.read_sql_table("types", conn, parse_dates={"IntDateCol": {"unit": "s"}})
    assert issubclass(df.IntDateCol.dtype.type, np.datetime64)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_datetime(conn, request):
    conn_name = conn
    conn = request.getfixturevalue(conn)
    df = DataFrame(
        {"A": date_range("2013-01-01 09:00:00", periods=3), "B": np.arange(3.0)}
    )
    assert df.to_sql(name="test_datetime", con=conn) == 3

    # with read_table -> type information from schema used
    result = sql.read_sql_table("test_datetime", conn)
    result = result.drop("index", axis=1)
    tm.assert_frame_equal(result, df)

    # with read_sql -> no type information -> sqlite has no native
    result = sql.read_sql_query("SELECT * FROM test_datetime", conn)
    result = result.drop("index", axis=1)
    if "sqlite" in conn_name:
        assert isinstance(result.loc[0, "A"], str)
        result["A"] = to_datetime(result["A"])
        tm.assert_frame_equal(result, df)
    else:
        tm.assert_frame_equal(result, df)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_datetime_NaT(conn, request):
    conn_name = conn
    conn = request.getfixturevalue(conn)
    df = DataFrame(
        {"A": date_range("2013-01-01 09:00:00", periods=3), "B": np.arange(3.0)}
    )
    df.loc[1, "A"] = np.nan
    assert df.to_sql(name="test_datetime", con=conn, index=False) == 3

    # with read_table -> type information from schema used
    result = sql.read_sql_table("test_datetime", conn)
    tm.assert_frame_equal(result, df)

    # with read_sql -> no type information -> sqlite has no native
    result = sql.read_sql_query("SELECT * FROM test_datetime", conn)
    if "sqlite" in conn_name:
        assert isinstance(result.loc[0, "A"], str)
        result["A"] = to_datetime(result["A"], errors="coerce")
        tm.assert_frame_equal(result, df)
    else:
        tm.assert_frame_equal(result, df)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_datetime_date(conn, request):
    # test support for datetime.date
    conn = request.getfixturevalue(conn)
    df = DataFrame([date(2014, 1, 1), date(2014, 1, 2)], columns=["a"])
    assert df.to_sql(name="test_date", con=conn, index=False) == 2
    res = read_sql_table("test_date", conn)
    result = res["a"]
    expected = to_datetime(df["a"])
    # comes back as datetime64
    tm.assert_series_equal(result, expected)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_datetime_time(conn, request, sqlite_buildin):
    # test support for datetime.time
    conn_name = conn
    conn = request.getfixturevalue(conn)
    df = DataFrame([time(9, 0, 0), time(9, 1, 30)], columns=["a"])
    assert df.to_sql(name="test_time", con=conn, index=False) == 2
    res = read_sql_table("test_time", conn)
    tm.assert_frame_equal(res, df)

    # GH8341
    # first, use the fallback to have the sqlite adapter put in place
    sqlite_conn = sqlite_buildin
    assert sql.to_sql(df, "test_time2", sqlite_conn, index=False) == 2
    res = sql.read_sql_query("SELECT * FROM test_time2", sqlite_conn)
    ref = df.map(lambda _: _.strftime("%H:%M:%S.%f"))
    tm.assert_frame_equal(ref, res)  # check if adapter is in place
    # then test if sqlalchemy is unaffected by the sqlite adapter
    assert sql.to_sql(df, "test_time3", conn, index=False) == 2
    if "sqlite" in conn_name:
        res = sql.read_sql_query("SELECT * FROM test_time3", conn)
        ref = df.map(lambda _: _.strftime("%H:%M:%S.%f"))
        tm.assert_frame_equal(ref, res)
    res = sql.read_sql_table("test_time3", conn)
    tm.assert_frame_equal(df, res)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_mixed_dtype_insert(conn, request):
    # see GH6509
    conn = request.getfixturevalue(conn)
    s1 = Series(2**25 + 1, dtype=np.int32)
    s2 = Series(0.0, dtype=np.float32)
    df = DataFrame({"s1": s1, "s2": s2})

    # write and read again
    assert df.to_sql(name="test_read_write", con=conn, index=False) == 1
    df2 = sql.read_sql_table("test_read_write", conn)

    tm.assert_frame_equal(df, df2, check_dtype=False, check_exact=True)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_nan_numeric(conn, request):
    # NaNs in numeric float column
    conn = request.getfixturevalue(conn)
    df = DataFrame({"A": [0, 1, 2], "B": [0.2, np.nan, 5.6]})
    assert df.to_sql(name="test_nan", con=conn, index=False) == 3

    # with read_table
    result = sql.read_sql_table("test_nan", conn)
    tm.assert_frame_equal(result, df)

    # with read_sql
    result = sql.read_sql_query("SELECT * FROM test_nan", conn)
    tm.assert_frame_equal(result, df)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_nan_fullcolumn(conn, request):
    # full NaN column (numeric float column)
    conn = request.getfixturevalue(conn)
    df = DataFrame({"A": [0, 1, 2], "B": [np.nan, np.nan, np.nan]})
    assert df.to_sql(name="test_nan", con=conn, index=False) == 3

    # with read_table
    result = sql.read_sql_table("test_nan", conn)
    tm.assert_frame_equal(result, df)

    # with read_sql -> not type info from table -> stays None
    df["B"] = df["B"].astype("object")
    df["B"] = None
    result = sql.read_sql_query("SELECT * FROM test_nan", conn)
    tm.assert_frame_equal(result, df)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_nan_string(conn, request):
    # NaNs in string column
    conn = request.getfixturevalue(conn)
    df = DataFrame({"A": [0, 1, 2], "B": ["a", "b", np.nan]})
    assert df.to_sql(name="test_nan", con=conn, index=False) == 3

    # NaNs are coming back as None
    df.loc[2, "B"] = None

    # with read_table
    result = sql.read_sql_table("test_nan", conn)
    tm.assert_frame_equal(result, df)

    # with read_sql
    result = sql.read_sql_query("SELECT * FROM test_nan", conn)
    tm.assert_frame_equal(result, df)


@pytest.mark.parametrize("conn", all_connectable)
def test_to_sql_save_index(conn, request):
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(
                reason="ADBC implementation does not create index", strict=True
            )
        )
    conn_name = conn
    conn = request.getfixturevalue(conn)
    df = DataFrame.from_records(
        [(1, 2.1, "line1"), (2, 1.5, "line2")], columns=["A", "B", "C"], index=["A"]
    )

    tbl_name = "test_to_sql_saves_index"
    with pandasSQL_builder(conn) as pandasSQL:
        with pandasSQL.run_transaction():
            assert pandasSQL.to_sql(df, tbl_name) == 2

    if conn_name in {"sqlite_buildin", "sqlite_str"}:
        ixs = sql.read_sql_query(
            "SELECT * FROM sqlite_master WHERE type = 'index' "
            f"AND tbl_name = '{tbl_name}'",
            conn,
        )
        ix_cols = []
        for ix_name in ixs.name:
            ix_info = sql.read_sql_query(f"PRAGMA index_info({ix_name})", conn)
            ix_cols.append(ix_info.name.tolist())
    else:
        from sqlalchemy import inspect

        insp = inspect(conn)

        ixs = insp.get_indexes(tbl_name)
        ix_cols = [i["column_names"] for i in ixs]

    assert ix_cols == [["A"]]


@pytest.mark.parametrize("conn", all_connectable)
def test_transactions(conn, request):
    conn_name = conn
    conn = request.getfixturevalue(conn)

    stmt = "CREATE TABLE test_trans (A INT, B TEXT)"
    if conn_name != "sqlite_buildin" and "adbc" not in conn_name:
        from sqlalchemy import text

        stmt = text(stmt)

    with pandasSQL_builder(conn) as pandasSQL:
        with pandasSQL.run_transaction() as trans:
            trans.execute(stmt)


@pytest.mark.parametrize("conn", all_connectable)
def test_transaction_rollback(conn, request):
    conn_name = conn
    conn = request.getfixturevalue(conn)
    with pandasSQL_builder(conn) as pandasSQL:
        with pandasSQL.run_transaction() as trans:
            stmt = "CREATE TABLE test_trans (A INT, B TEXT)"
            if "adbc" in conn_name or isinstance(pandasSQL, SQLiteDatabase):
                trans.execute(stmt)
            else:
                from sqlalchemy import text

                stmt = text(stmt)
                trans.execute(stmt)

        class DummyException(Exception):
            pass

        # Make sure when transaction is rolled back, no rows get inserted
        ins_sql = "INSERT INTO test_trans (A,B) VALUES (1, 'blah')"
        if isinstance(pandasSQL, SQLDatabase):
            from sqlalchemy import text

            ins_sql = text(ins_sql)
        try:
            with pandasSQL.run_transaction() as trans:
                trans.execute(ins_sql)
                raise DummyException("error")
        except DummyException:
            # ignore raised exception
            pass
        with pandasSQL.run_transaction():
            res = pandasSQL.read_query("SELECT * FROM test_trans")
        assert len(res) == 0

        # Make sure when transaction is committed, rows do get inserted
        with pandasSQL.run_transaction() as trans:
            trans.execute(ins_sql)
            res2 = pandasSQL.read_query("SELECT * FROM test_trans")
        assert len(res2) == 1


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_get_schema_create_table(conn, request, test_frame3):
    # Use a dataframe without a bool column, since MySQL converts bool to
    # TINYINT (which read_sql_table returns as an int and causes a dtype
    # mismatch)
    if conn == "sqlite_str":
        request.applymarker(
            pytest.mark.xfail(reason="test does not support sqlite_str fixture")
        )

    conn = request.getfixturevalue(conn)

    from sqlalchemy import text
    from sqlalchemy.engine import Engine

    tbl = "test_get_schema_create_table"
    create_sql = sql.get_schema(test_frame3, tbl, con=conn)
    blank_test_df = test_frame3.iloc[:0]

    create_sql = text(create_sql)
    if isinstance(conn, Engine):
        with conn.connect() as newcon:
            with newcon.begin():
                newcon.execute(create_sql)
    else:
        conn.execute(create_sql)
    returned_df = sql.read_sql_table(tbl, conn)
    tm.assert_frame_equal(returned_df, blank_test_df, check_index_type=False)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_dtype(conn, request):
    if conn == "sqlite_str":
        pytest.skip("sqlite_str has no inspection system")

    conn = request.getfixturevalue(conn)

    from sqlalchemy import (
        TEXT,
        String,
    )
    from sqlalchemy.schema import MetaData

    cols = ["A", "B"]
    data = [(0.8, True), (0.9, None)]
    df = DataFrame(data, columns=cols)
    assert df.to_sql(name="dtype_test", con=conn) == 2
    assert df.to_sql(name="dtype_test2", con=conn, dtype={"B": TEXT}) == 2
    meta = MetaData()
    meta.reflect(bind=conn)
    sqltype = meta.tables["dtype_test2"].columns["B"].type
    assert isinstance(sqltype, TEXT)
    msg = "The type of B is not a SQLAlchemy type"
    with pytest.raises(ValueError, match=msg):
        df.to_sql(name="error", con=conn, dtype={"B": str})

    # GH9083
    assert df.to_sql(name="dtype_test3", con=conn, dtype={"B": String(10)}) == 2
    meta.reflect(bind=conn)
    sqltype = meta.tables["dtype_test3"].columns["B"].type
    assert isinstance(sqltype, String)
    assert sqltype.length == 10

    # single dtype
    assert df.to_sql(name="single_dtype_test", con=conn, dtype=TEXT) == 2
    meta.reflect(bind=conn)
    sqltypea = meta.tables["single_dtype_test"].columns["A"].type
    sqltypeb = meta.tables["single_dtype_test"].columns["B"].type
    assert isinstance(sqltypea, TEXT)
    assert isinstance(sqltypeb, TEXT)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_notna_dtype(conn, request):
    if conn == "sqlite_str":
        pytest.skip("sqlite_str has no inspection system")

    conn_name = conn
    conn = request.getfixturevalue(conn)

    from sqlalchemy import (
        Boolean,
        DateTime,
        Float,
        Integer,
    )
    from sqlalchemy.schema import MetaData

    cols = {
        "Bool": Series([True, None]),
        "Date": Series([datetime(2012, 5, 1), None]),
        "Int": Series([1, None], dtype="object"),
        "Float": Series([1.1, None]),
    }
    df = DataFrame(cols)

    tbl = "notna_dtype_test"
    assert df.to_sql(name=tbl, con=conn) == 2
    _ = sql.read_sql_table(tbl, conn)
    meta = MetaData()
    meta.reflect(bind=conn)
    my_type = Integer if "mysql" in conn_name else Boolean
    col_dict = meta.tables[tbl].columns
    assert isinstance(col_dict["Bool"].type, my_type)
    assert isinstance(col_dict["Date"].type, DateTime)
    assert isinstance(col_dict["Int"].type, Integer)
    assert isinstance(col_dict["Float"].type, Float)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_double_precision(conn, request):
    if conn == "sqlite_str":
        pytest.skip("sqlite_str has no inspection system")

    conn = request.getfixturevalue(conn)

    from sqlalchemy import (
        BigInteger,
        Float,
        Integer,
    )
    from sqlalchemy.schema import MetaData

    V = 1.23456789101112131415

    df = DataFrame(
        {
            "f32": Series([V], dtype="float32"),
            "f64": Series([V], dtype="float64"),
            "f64_as_f32": Series([V], dtype="float64"),
            "i32": Series([5], dtype="int32"),
            "i64": Series([5], dtype="int64"),
        }
    )

    assert (
        df.to_sql(
            name="test_dtypes",
            con=conn,
            index=False,
            if_exists="replace",
            dtype={"f64_as_f32": Float(precision=23)},
        )
        == 1
    )
    res = sql.read_sql_table("test_dtypes", conn)

    # check precision of float64
    assert np.round(df["f64"].iloc[0], 14) == np.round(res["f64"].iloc[0], 14)

    # check sql types
    meta = MetaData()
    meta.reflect(bind=conn)
    col_dict = meta.tables["test_dtypes"].columns
    assert str(col_dict["f32"].type) == str(col_dict["f64_as_f32"].type)
    assert isinstance(col_dict["f32"].type, Float)
    assert isinstance(col_dict["f64"].type, Float)
    assert isinstance(col_dict["i32"].type, Integer)
    assert isinstance(col_dict["i64"].type, BigInteger)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_connectable_issue_example(conn, request):
    conn = request.getfixturevalue(conn)

    # This tests the example raised in issue
    # https://github.com/pandas-dev/pandas/issues/10104
    from sqlalchemy.engine import Engine

    def test_select(connection):
        query = "SELECT test_foo_data FROM test_foo_data"
        return sql.read_sql_query(query, con=connection)

    def test_append(connection, data):
        data.to_sql(name="test_foo_data", con=connection, if_exists="append")

    def test_connectable(conn):
        # https://github.com/sqlalchemy/sqlalchemy/commit/
        # 00b5c10846e800304caa86549ab9da373b42fa5d#r48323973
        foo_data = test_select(conn)
        test_append(conn, foo_data)

    def main(connectable):
        if isinstance(connectable, Engine):
            with connectable.connect() as conn:
                with conn.begin():
                    test_connectable(conn)
        else:
            test_connectable(connectable)

    assert (
        DataFrame({"test_foo_data": [0, 1, 2]}).to_sql(name="test_foo_data", con=conn)
        == 3
    )
    main(conn)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
@pytest.mark.parametrize(
    "input",
    [{"foo": [np.inf]}, {"foo": [-np.inf]}, {"foo": [-np.inf], "infe0": ["bar"]}],
)
def test_to_sql_with_negative_npinf(conn, request, input):
    # GH 34431

    df = DataFrame(input)
    conn_name = conn
    conn = request.getfixturevalue(conn)

    if "mysql" in conn_name:
        # GH 36465
        # The input {"foo": [-np.inf], "infe0": ["bar"]} does not raise any error
        # for pymysql version >= 0.10
        # TODO(GH#36465): remove this version check after GH 36465 is fixed
        pymysql = pytest.importorskip("pymysql")

        if Version(pymysql.__version__) < Version("1.0.3") and "infe0" in df.columns:
            mark = pytest.mark.xfail(reason="GH 36465")
            request.applymarker(mark)

        msg = "inf cannot be used with MySQL"
        with pytest.raises(ValueError, match=msg):
            df.to_sql(name="foobar", con=conn, index=False)
    else:
        assert df.to_sql(name="foobar", con=conn, index=False) == 1
        res = sql.read_sql_table("foobar", conn)
        tm.assert_equal(df, res)


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_temporary_table(conn, request):
    if conn == "sqlite_str":
        pytest.skip("test does not work with str connection")

    conn = request.getfixturevalue(conn)

    from sqlalchemy import (
        Column,
        Integer,
        Unicode,
        select,
    )
    from sqlalchemy.orm import (
        Session,
        declarative_base,
    )

    test_data = "Hello, World!"
    expected = DataFrame({"spam": [test_data]})
    Base = declarative_base()

    class Temporary(Base):
        __tablename__ = "temp_test"
        __table_args__ = {"prefixes": ["TEMPORARY"]}
        id = Column(Integer, primary_key=True)
        spam = Column(Unicode(30), nullable=False)

    with Session(conn) as session:
        with session.begin():
            conn = session.connection()
            Temporary.__table__.create(conn)
            session.add(Temporary(spam=test_data))
            session.flush()
            df = sql.read_sql_query(sql=select(Temporary.spam), con=conn)
    tm.assert_frame_equal(df, expected)


@pytest.mark.parametrize("conn", all_connectable)
def test_invalid_engine(conn, request, test_frame1):
    if conn == "sqlite_buildin" or "adbc" in conn:
        request.applymarker(
            pytest.mark.xfail(
                reason="SQLiteDatabase/ADBCDatabase does not raise for bad engine"
            )
        )

    conn = request.getfixturevalue(conn)
    msg = "engine must be one of 'auto', 'sqlalchemy'"
    with pandasSQL_builder(conn) as pandasSQL:
        with pytest.raises(ValueError, match=msg):
            pandasSQL.to_sql(test_frame1, "test_frame1", engine="bad_engine")


@pytest.mark.parametrize("conn", all_connectable)
def test_to_sql_with_sql_engine(conn, request, test_frame1):
    """`to_sql` with the `engine` param"""
    # mostly copied from this class's `_to_sql()` method
    conn = request.getfixturevalue(conn)
    with pandasSQL_builder(conn) as pandasSQL:
        with pandasSQL.run_transaction():
            assert pandasSQL.to_sql(test_frame1, "test_frame1", engine="auto") == 4
            assert pandasSQL.has_table("test_frame1")

    num_entries = len(test_frame1)
    num_rows = count_rows(conn, "test_frame1")
    assert num_rows == num_entries


@pytest.mark.parametrize("conn", sqlalchemy_connectable)
def test_options_sqlalchemy(conn, request, test_frame1):
    # use the set option
    conn = request.getfixturevalue(conn)
    with pd.option_context("io.sql.engine", "sqlalchemy"):
        with pandasSQL_builder(conn) as pandasSQL:
            with pandasSQL.run_transaction():
                assert pandasSQL.to_sql(test_frame1, "test_frame1") == 4
                assert pandasSQL.has_table("test_frame1")

        num_entries = len(test_frame1)
        num_rows = count_rows(conn, "test_frame1")
        assert num_rows == num_entries


@pytest.mark.parametrize("conn", all_connectable)
def test_options_auto(conn, request, test_frame1):
    # use the set option
    conn = request.getfixturevalue(conn)
    with pd.option_context("io.sql.engine", "auto"):
        with pandasSQL_builder(conn) as pandasSQL:
            with pandasSQL.run_transaction():
                assert pandasSQL.to_sql(test_frame1, "test_frame1") == 4
                assert pandasSQL.has_table("test_frame1")

        num_entries = len(test_frame1)
        num_rows = count_rows(conn, "test_frame1")
        assert num_rows == num_entries


def test_options_get_engine():
    pytest.importorskip("sqlalchemy")
    assert isinstance(get_engine("sqlalchemy"), SQLAlchemyEngine)

    with pd.option_context("io.sql.engine", "sqlalchemy"):
        assert isinstance(get_engine("auto"), SQLAlchemyEngine)
        assert isinstance(get_engine("sqlalchemy"), SQLAlchemyEngine)

    with pd.option_context("io.sql.engine", "auto"):
        assert isinstance(get_engine("auto"), SQLAlchemyEngine)
        assert isinstance(get_engine("sqlalchemy"), SQLAlchemyEngine)


def test_get_engine_auto_error_message():
    # Expect different error messages from get_engine(engine="auto")
    # if engines aren't installed vs. are installed but bad version
    pass
    # TODO(GH#36893) fill this in when we add more engines


@pytest.mark.parametrize("conn", all_connectable)
@pytest.mark.parametrize("func", ["read_sql", "read_sql_query"])
def test_read_sql_dtype_backend(
    conn,
    request,
    string_storage,
    func,
    dtype_backend,
    dtype_backend_data,
    dtype_backend_expected,
):
    # GH#50048
    conn_name = conn
    conn = request.getfixturevalue(conn)
    table = "test"
    df = dtype_backend_data
    df.to_sql(name=table, con=conn, index=False, if_exists="replace")

    with pd.option_context("mode.string_storage", string_storage):
        result = getattr(pd, func)(
            f"Select * from {table}", conn, dtype_backend=dtype_backend
        )
        expected = dtype_backend_expected(string_storage, dtype_backend, conn_name)

    tm.assert_frame_equal(result, expected)

    if "adbc" in conn_name:
        # adbc does not support chunksize argument
        request.applymarker(
            pytest.mark.xfail(reason="adbc does not support chunksize argument")
        )

    with pd.option_context("mode.string_storage", string_storage):
        iterator = getattr(pd, func)(
            f"Select * from {table}",
            con=conn,
            dtype_backend=dtype_backend,
            chunksize=3,
        )
        expected = dtype_backend_expected(string_storage, dtype_backend, conn_name)
        for result in iterator:
            tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("conn", all_connectable)
@pytest.mark.parametrize("func", ["read_sql", "read_sql_table"])
def test_read_sql_dtype_backend_table(
    conn,
    request,
    string_storage,
    func,
    dtype_backend,
    dtype_backend_data,
    dtype_backend_expected,
):
    if "sqlite" in conn and "adbc" not in conn:
        request.applymarker(
            pytest.mark.xfail(
                reason=(
                    "SQLite actually returns proper boolean values via "
                    "read_sql_table, but before pytest refactor was skipped"
                )
            )
        )
    # GH#50048
    conn_name = conn
    conn = request.getfixturevalue(conn)
    table = "test"
    df = dtype_backend_data
    df.to_sql(name=table, con=conn, index=False, if_exists="replace")

    with pd.option_context("mode.string_storage", string_storage):
        result = getattr(pd, func)(table, conn, dtype_backend=dtype_backend)
        expected = dtype_backend_expected(string_storage, dtype_backend, conn_name)
    tm.assert_frame_equal(result, expected)

    if "adbc" in conn_name:
        # adbc does not support chunksize argument
        return

    with pd.option_context("mode.string_storage", string_storage):
        iterator = getattr(pd, func)(
            table,
            conn,
            dtype_backend=dtype_backend,
            chunksize=3,
        )
        expected = dtype_backend_expected(string_storage, dtype_backend, conn_name)
        for result in iterator:
            tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("conn", all_connectable)
@pytest.mark.parametrize("func", ["read_sql", "read_sql_table", "read_sql_query"])
def test_read_sql_invalid_dtype_backend_table(conn, request, func, dtype_backend_data):
    conn = request.getfixturevalue(conn)
    table = "test"
    df = dtype_backend_data
    df.to_sql(name=table, con=conn, index=False, if_exists="replace")

    msg = (
        "dtype_backend numpy is invalid, only 'numpy_nullable' and "
        "'pyarrow' are allowed."
    )
    with pytest.raises(ValueError, match=msg):
        getattr(pd, func)(table, conn, dtype_backend="numpy")


@pytest.fixture
def dtype_backend_data() -> DataFrame:
    return DataFrame(
        {
            "a": Series([1, np.nan, 3], dtype="Int64"),
            "b": Series([1, 2, 3], dtype="Int64"),
            "c": Series([1.5, np.nan, 2.5], dtype="Float64"),
            "d": Series([1.5, 2.0, 2.5], dtype="Float64"),
            "e": [True, False, None],
            "f": [True, False, True],
            "g": ["a", "b", "c"],
            "h": ["a", "b", None],
        }
    )


@pytest.fixture
def dtype_backend_expected():
    def func(string_storage, dtype_backend, conn_name) -> DataFrame:
        string_dtype: pd.StringDtype | pd.ArrowDtype
        if dtype_backend == "pyarrow":
            pa = pytest.importorskip("pyarrow")
            string_dtype = pd.ArrowDtype(pa.string())
        else:
            string_dtype = pd.StringDtype(string_storage)

        df = DataFrame(
            {
                "a": Series([1, np.nan, 3], dtype="Int64"),
                "b": Series([1, 2, 3], dtype="Int64"),
                "c": Series([1.5, np.nan, 2.5], dtype="Float64"),
                "d": Series([1.5, 2.0, 2.5], dtype="Float64"),
                "e": Series([True, False, pd.NA], dtype="boolean"),
                "f": Series([True, False, True], dtype="boolean"),
                "g": Series(["a", "b", "c"], dtype=string_dtype),
                "h": Series(["a", "b", None], dtype=string_dtype),
            }
        )
        if dtype_backend == "pyarrow":
            pa = pytest.importorskip("pyarrow")

            from pandas.arrays import ArrowExtensionArray

            df = DataFrame(
                {
                    col: ArrowExtensionArray(pa.array(df[col], from_pandas=True))
                    for col in df.columns
                }
            )

        if "mysql" in conn_name or "sqlite" in conn_name:
            if dtype_backend == "numpy_nullable":
                df = df.astype({"e": "Int64", "f": "Int64"})
            else:
                df = df.astype({"e": "int64[pyarrow]", "f": "int64[pyarrow]"})

        return df

    return func


@pytest.mark.parametrize("conn", all_connectable)
def test_chunksize_empty_dtypes(conn, request):
    # GH#50245
    if "adbc" in conn:
        request.node.add_marker(
            pytest.mark.xfail(reason="chunksize argument NotImplemented with ADBC")
        )
    conn = request.getfixturevalue(conn)
    dtypes = {"a": "int64", "b": "object"}
    df = DataFrame(columns=["a", "b"]).astype(dtypes)
    expected = df.copy()
    df.to_sql(name="test", con=conn, index=False, if_exists="replace")

    for result in read_sql_query(
        "SELECT * FROM test",
        conn,
        dtype=dtypes,
        chunksize=1,
    ):
        tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("conn", all_connectable)
@pytest.mark.parametrize("dtype_backend", [lib.no_default, "numpy_nullable"])
@pytest.mark.parametrize("func", ["read_sql", "read_sql_query"])
def test_read_sql_dtype(conn, request, func, dtype_backend):
    # GH#50797
    conn = request.getfixturevalue(conn)
    table = "test"
    df = DataFrame({"a": [1, 2, 3], "b": 5})
    df.to_sql(name=table, con=conn, index=False, if_exists="replace")

    result = getattr(pd, func)(
        f"Select * from {table}",
        conn,
        dtype={"a": np.float64},
        dtype_backend=dtype_backend,
    )
    expected = DataFrame(
        {
            "a": Series([1, 2, 3], dtype=np.float64),
            "b": Series(
                [5, 5, 5],
                dtype="int64" if not dtype_backend == "numpy_nullable" else "Int64",
            ),
        }
    )
    tm.assert_frame_equal(result, expected)


def test_keyword_deprecation(sqlite_engine):
    conn = sqlite_engine
    # GH 54397
    msg = (
        "Starting with pandas version 3.0 all arguments of to_sql except for the "
        "arguments 'name' and 'con' will be keyword-only."
    )
    df = DataFrame([{"A": 1, "B": 2, "C": 3}, {"A": 1, "B": 2, "C": 3}])
    df.to_sql("example", conn)

    with tm.assert_produces_warning(FutureWarning, match=msg):
        df.to_sql("example", conn, None, if_exists="replace")


def test_bigint_warning(sqlite_engine):
    conn = sqlite_engine
    # test no warning for BIGINT (to support int64) is raised (GH7433)
    df = DataFrame({"a": [1, 2]}, dtype="int64")
    assert df.to_sql(name="test_bigintwarning", con=conn, index=False) == 2

    with tm.assert_produces_warning(None):
        sql.read_sql_table("test_bigintwarning", conn)


def test_valueerror_exception(sqlite_engine):
    conn = sqlite_engine
    df = DataFrame({"col1": [1, 2], "col2": [3, 4]})
    with pytest.raises(ValueError, match="Empty table name specified"):
        df.to_sql(name="", con=conn, if_exists="replace", index=False)


def test_row_object_is_named_tuple(sqlite_engine):
    conn = sqlite_engine
    # GH 40682
    # Test for the is_named_tuple() function
    # Placed here due to its usage of sqlalchemy

    from sqlalchemy import (
        Column,
        Integer,
        String,
    )
    from sqlalchemy.orm import (
        declarative_base,
        sessionmaker,
    )

    BaseModel = declarative_base()

    class Test(BaseModel):
        __tablename__ = "test_frame"
        id = Column(Integer, primary_key=True)
        string_column = Column(String(50))

    with conn.begin():
        BaseModel.metadata.create_all(conn)
    Session = sessionmaker(bind=conn)
    with Session() as session:
        df = DataFrame({"id": [0, 1], "string_column": ["hello", "world"]})
        assert (
            df.to_sql(name="test_frame", con=conn, index=False, if_exists="replace")
            == 2
        )
        session.commit()
        test_query = session.query(Test.id, Test.string_column)
        df = DataFrame(test_query)

    assert list(df.columns) == ["id", "string_column"]


def test_read_sql_string_inference(sqlite_engine):
    conn = sqlite_engine
    # GH#54430
    table = "test"
    df = DataFrame({"a": ["x", "y"]})
    df.to_sql(table, con=conn, index=False, if_exists="replace")

    with pd.option_context("future.infer_string", True):
        result = read_sql_table(table, conn)

    dtype = pd.StringDtype(na_value=np.nan)
    expected = DataFrame(
        {"a": ["x", "y"]}, dtype=dtype, columns=Index(["a"], dtype=dtype)
    )

    tm.assert_frame_equal(result, expected)


def test_roundtripping_datetimes(sqlite_engine):
    conn = sqlite_engine
    # GH#54877
    df = DataFrame({"t": [datetime(2020, 12, 31, 12)]}, dtype="datetime64[ns]")
    df.to_sql("test", conn, if_exists="replace", index=False)
    result = pd.read_sql("select * from test", conn).iloc[0, 0]
    assert result == "2020-12-31 12:00:00.000000"


@pytest.fixture
def sqlite_builtin_detect_types():
    with contextlib.closing(
        sqlite3.connect(":memory:", detect_types=sqlite3.PARSE_DECLTYPES)
    ) as closing_conn:
        with closing_conn as conn:
            yield conn


def test_roundtripping_datetimes_detect_types(sqlite_builtin_detect_types):
    # https://github.com/pandas-dev/pandas/issues/55554
    conn = sqlite_builtin_detect_types
    df = DataFrame({"t": [datetime(2020, 12, 31, 12)]}, dtype="datetime64[ns]")
    df.to_sql("test", conn, if_exists="replace", index=False)
    result = pd.read_sql("select * from test", conn).iloc[0, 0]
    assert result == Timestamp("2020-12-31 12:00:00.000000")


@pytest.mark.db
def test_psycopg2_schema_support(postgresql_psycopg2_engine):
    conn = postgresql_psycopg2_engine

    # only test this for postgresql (schema's not supported in
    # mysql/sqlite)
    df = DataFrame({"col1": [1, 2], "col2": [0.1, 0.2], "col3": ["a", "n"]})

    # create a schema
    with conn.connect() as con:
        with con.begin():
            con.exec_driver_sql("DROP SCHEMA IF EXISTS other CASCADE;")
            con.exec_driver_sql("CREATE SCHEMA other;")

    # write dataframe to different schema's
    assert df.to_sql(name="test_schema_public", con=conn, index=False) == 2
    assert (
        df.to_sql(
            name="test_schema_public_explicit",
            con=conn,
            index=False,
            schema="public",
        )
        == 2
    )
    assert (
        df.to_sql(name="test_schema_other", con=conn, index=False, schema="other") == 2
    )

    # read dataframes back in
    res1 = sql.read_sql_table("test_schema_public", conn)
    tm.assert_frame_equal(df, res1)
    res2 = sql.read_sql_table("test_schema_public_explicit", conn)
    tm.assert_frame_equal(df, res2)
    res3 = sql.read_sql_table("test_schema_public_explicit", conn, schema="public")
    tm.assert_frame_equal(df, res3)
    res4 = sql.read_sql_table("test_schema_other", conn, schema="other")
    tm.assert_frame_equal(df, res4)
    msg = "Table test_schema_other not found"
    with pytest.raises(ValueError, match=msg):
        sql.read_sql_table("test_schema_other", conn, schema="public")

    # different if_exists options

    # create a schema
    with conn.connect() as con:
        with con.begin():
            con.exec_driver_sql("DROP SCHEMA IF EXISTS other CASCADE;")
            con.exec_driver_sql("CREATE SCHEMA other;")

    # write dataframe with different if_exists options
    assert (
        df.to_sql(name="test_schema_other", con=conn, schema="other", index=False) == 2
    )
    df.to_sql(
        name="test_schema_other",
        con=conn,
        schema="other",
        index=False,
        if_exists="replace",
    )
    assert (
        df.to_sql(
            name="test_schema_other",
            con=conn,
            schema="other",
            index=False,
            if_exists="append",
        )
        == 2
    )
    res = sql.read_sql_table("test_schema_other", conn, schema="other")
    tm.assert_frame_equal(concat([df, df], ignore_index=True), res)


@pytest.mark.db
def test_self_join_date_columns(postgresql_psycopg2_engine):
    # GH 44421
    conn = postgresql_psycopg2_engine
    from sqlalchemy.sql import text

    create_table = text(
        """
    CREATE TABLE person
    (
        id serial constraint person_pkey primary key,
        created_dt timestamp with time zone
    );

    INSERT INTO person
        VALUES (1, '2021-01-01T00:00:00Z');
    """
    )
    with conn.connect() as con:
        with con.begin():
            con.execute(create_table)

    sql_query = (
        'SELECT * FROM "person" AS p1 INNER JOIN "person" AS p2 ON p1.id = p2.id;'
    )
    result = pd.read_sql(sql_query, conn)
    expected = DataFrame(
        [[1, Timestamp("2021", tz="UTC")] * 2], columns=["id", "created_dt"] * 2
    )
    tm.assert_frame_equal(result, expected)

    # Cleanup
    with sql.SQLDatabase(conn, need_transaction=True) as pandasSQL:
        pandasSQL.drop_table("person")


def test_create_and_drop_table(sqlite_engine):
    conn = sqlite_engine
    temp_frame = DataFrame({"one": [1.0, 2.0, 3.0, 4.0], "two": [4.0, 3.0, 2.0, 1.0]})
    with sql.SQLDatabase(conn) as pandasSQL:
        with pandasSQL.run_transaction():
            assert pandasSQL.to_sql(temp_frame, "drop_test_frame") == 4

        assert pandasSQL.has_table("drop_test_frame")

        with pandasSQL.run_transaction():
            pandasSQL.drop_table("drop_test_frame")

        assert not pandasSQL.has_table("drop_test_frame")


def test_sqlite_datetime_date(sqlite_buildin):
    conn = sqlite_buildin
    df = DataFrame([date(2014, 1, 1), date(2014, 1, 2)], columns=["a"])
    assert df.to_sql(name="test_date", con=conn, index=False) == 2
    res = read_sql_query("SELECT * FROM test_date", conn)
    # comes back as strings
    tm.assert_frame_equal(res, df.astype(str))


@pytest.mark.parametrize("tz_aware", [False, True])
def test_sqlite_datetime_time(tz_aware, sqlite_buildin):
    conn = sqlite_buildin
    # test support for datetime.time, GH #8341
    if not tz_aware:
        tz_times = [time(9, 0, 0), time(9, 1, 30)]
    else:
        tz_dt = date_range("2013-01-01 09:00:00", periods=2, tz="US/Pacific")
        tz_times = Series(tz_dt.to_pydatetime()).map(lambda dt: dt.timetz())

    df = DataFrame(tz_times, columns=["a"])

    assert df.to_sql(name="test_time", con=conn, index=False) == 2
    res = read_sql_query("SELECT * FROM test_time", conn)
    # comes back as strings
    expected = df.map(lambda _: _.strftime("%H:%M:%S.%f"))
    tm.assert_frame_equal(res, expected)


def get_sqlite_column_type(conn, table, column):
    recs = conn.execute(f"PRAGMA table_info({table})")
    for cid, name, ctype, not_null, default, pk in recs:
        if name == column:
            return ctype
    raise ValueError(f"Table {table}, column {column} not found")


def test_sqlite_test_dtype(sqlite_buildin):
    conn = sqlite_buildin
    cols = ["A", "B"]
    data = [(0.8, True), (0.9, None)]
    df = DataFrame(data, columns=cols)
    assert df.to_sql(name="dtype_test", con=conn) == 2
    assert df.to_sql(name="dtype_test2", con=conn, dtype={"B": "STRING"}) == 2

    # sqlite stores Boolean values as INTEGER
    assert get_sqlite_column_type(conn, "dtype_test", "B") == "INTEGER"

    assert get_sqlite_column_type(conn, "dtype_test2", "B") == "STRING"
    msg = r"B \(<class 'bool'>\) not a string"
    with pytest.raises(ValueError, match=msg):
        df.to_sql(name="error", con=conn, dtype={"B": bool})

    # single dtype
    assert df.to_sql(name="single_dtype_test", con=conn, dtype="STRING") == 2
    assert get_sqlite_column_type(conn, "single_dtype_test", "A") == "STRING"
    assert get_sqlite_column_type(conn, "single_dtype_test", "B") == "STRING"


def test_sqlite_notna_dtype(sqlite_buildin):
    conn = sqlite_buildin
    cols = {
        "Bool": Series([True, None]),
        "Date": Series([datetime(2012, 5, 1), None]),
        "Int": Series([1, None], dtype="object"),
        "Float": Series([1.1, None]),
    }
    df = DataFrame(cols)

    tbl = "notna_dtype_test"
    assert df.to_sql(name=tbl, con=conn) == 2

    assert get_sqlite_column_type(conn, tbl, "Bool") == "INTEGER"
    assert get_sqlite_column_type(conn, tbl, "Date") == "TIMESTAMP"
    assert get_sqlite_column_type(conn, tbl, "Int") == "INTEGER"
    assert get_sqlite_column_type(conn, tbl, "Float") == "REAL"


def test_sqlite_illegal_names(sqlite_buildin):
    # For sqlite, these should work fine
    conn = sqlite_buildin
    df = DataFrame([[1, 2], [3, 4]], columns=["a", "b"])

    msg = "Empty table or column name specified"
    with pytest.raises(ValueError, match=msg):
        df.to_sql(name="", con=conn)

    for ndx, weird_name in enumerate(
        [
            "test_weird_name]",
            "test_weird_name[",
            "test_weird_name`",
            'test_weird_name"',
            "test_weird_name'",
            "_b.test_weird_name_01-30",
            '"_b.test_weird_name_01-30"',
            "99beginswithnumber",
            "12345",
            "\xe9",
        ]
    ):
        assert df.to_sql(name=weird_name, con=conn) == 2
        sql.table_exists(weird_name, conn)

        df2 = DataFrame([[1, 2], [3, 4]], columns=["a", weird_name])
        c_tbl = f"test_weird_col_name{ndx:d}"
        assert df2.to_sql(name=c_tbl, con=conn) == 2
        sql.table_exists(c_tbl, conn)


def format_query(sql, *args):
    _formatters = {
        datetime: "'{}'".format,
        str: "'{}'".format,
        np.str_: "'{}'".format,
        bytes: "'{}'".format,
        float: "{:.8f}".format,
        int: "{:d}".format,
        type(None): lambda x: "NULL",
        np.float64: "{:.10f}".format,
        bool: "'{!s}'".format,
    }
    processed_args = []
    for arg in args:
        if isinstance(arg, float) and isna(arg):
            arg = None

        formatter = _formatters[type(arg)]
        processed_args.append(formatter(arg))

    return sql % tuple(processed_args)


def tquery(query, con=None):
    """Replace removed sql.tquery function"""
    with sql.pandasSQL_builder(con) as pandas_sql:
        res = pandas_sql.execute(query).fetchall()
    return None if res is None else list(res)


def test_xsqlite_basic(sqlite_buildin):
    frame = DataFrame(
        np.random.default_rng(2).standard_normal((10, 4)),
        columns=Index(list("ABCD")),
        index=date_range("2000-01-01", periods=10, freq="B"),
    )
    assert sql.to_sql(frame, name="test_table", con=sqlite_buildin, index=False) == 10
    result = sql.read_sql("select * from test_table", sqlite_buildin)

    # HACK! Change this once indexes are handled properly.
    result.index = frame.index

    expected = frame
    tm.assert_frame_equal(result, frame)

    frame["txt"] = ["a"] * len(frame)
    frame2 = frame.copy()
    new_idx = Index(np.arange(len(frame2)), dtype=np.int64) + 10
    frame2["Idx"] = new_idx.copy()
    assert sql.to_sql(frame2, name="test_table2", con=sqlite_buildin, index=False) == 10
    result = sql.read_sql("select * from test_table2", sqlite_buildin, index_col="Idx")
    expected = frame.copy()
    expected.index = new_idx
    expected.index.name = "Idx"
    tm.assert_frame_equal(expected, result)


def test_xsqlite_write_row_by_row(sqlite_buildin):
    frame = DataFrame(
        np.random.default_rng(2).standard_normal((10, 4)),
        columns=Index(list("ABCD")),
        index=date_range("2000-01-01", periods=10, freq="B"),
    )
    frame.iloc[0, 0] = np.nan
    create_sql = sql.get_schema(frame, "test")
    cur = sqlite_buildin.cursor()
    cur.execute(create_sql)

    ins = "INSERT INTO test VALUES (%s, %s, %s, %s)"
    for _, row in frame.iterrows():
        fmt_sql = format_query(ins, *row)
        tquery(fmt_sql, con=sqlite_buildin)

    sqlite_buildin.commit()

    result = sql.read_sql("select * from test", con=sqlite_buildin)
    result.index = frame.index
    tm.assert_frame_equal(result, frame, rtol=1e-3)


def test_xsqlite_execute(sqlite_buildin):
    frame = DataFrame(
        np.random.default_rng(2).standard_normal((10, 4)),
        columns=Index(list("ABCD")),
        index=date_range("2000-01-01", periods=10, freq="B"),
    )
    create_sql = sql.get_schema(frame, "test")
    cur = sqlite_buildin.cursor()
    cur.execute(create_sql)
    ins = "INSERT INTO test VALUES (?, ?, ?, ?)"

    row = frame.iloc[0]
    with sql.pandasSQL_builder(sqlite_buildin) as pandas_sql:
        pandas_sql.execute(ins, tuple(row))
    sqlite_buildin.commit()

    result = sql.read_sql("select * from test", sqlite_buildin)
    result.index = frame.index[:1]
    tm.assert_frame_equal(result, frame[:1])


def test_xsqlite_schema(sqlite_buildin):
    frame = DataFrame(
        np.random.default_rng(2).standard_normal((10, 4)),
        columns=Index(list("ABCD")),
        index=date_range("2000-01-01", periods=10, freq="B"),
    )
    create_sql = sql.get_schema(frame, "test")
    lines = create_sql.splitlines()
    for line in lines:
        tokens = line.split(" ")
        if len(tokens) == 2 and tokens[0] == "A":
            assert tokens[1] == "DATETIME"

    create_sql = sql.get_schema(frame, "test", keys=["A", "B"])
    lines = create_sql.splitlines()
    assert 'PRIMARY KEY ("A", "B")' in create_sql
    cur = sqlite_buildin.cursor()
    cur.execute(create_sql)


def test_xsqlite_execute_fail(sqlite_buildin):
    create_sql = """
    CREATE TABLE test
    (
    a TEXT,
    b TEXT,
    c REAL,
    PRIMARY KEY (a, b)
    );
    """
    cur = sqlite_buildin.cursor()
    cur.execute(create_sql)

    with sql.pandasSQL_builder(sqlite_buildin) as pandas_sql:
        pandas_sql.execute("INSERT INTO test VALUES('foo', 'bar', 1.234)")
        pandas_sql.execute("INSERT INTO test VALUES('foo', 'baz', 2.567)")

        with pytest.raises(sql.DatabaseError, match="Execution failed on sql"):
            pandas_sql.execute("INSERT INTO test VALUES('foo', 'bar', 7)")


def test_xsqlite_execute_closed_connection():
    create_sql = """
    CREATE TABLE test
    (
    a TEXT,
    b TEXT,
    c REAL,
    PRIMARY KEY (a, b)
    );
    """
    with contextlib.closing(sqlite3.connect(":memory:")) as conn:
        cur = conn.cursor()
        cur.execute(create_sql)

        with sql.pandasSQL_builder(conn) as pandas_sql:
            pandas_sql.execute("INSERT INTO test VALUES('foo', 'bar', 1.234)")

    msg = "Cannot operate on a closed database."
    with pytest.raises(sqlite3.ProgrammingError, match=msg):
        tquery("select * from test", con=conn)


def test_xsqlite_keyword_as_column_names(sqlite_buildin):
    df = DataFrame({"From": np.ones(5)})
    assert sql.to_sql(df, con=sqlite_buildin, name="testkeywords", index=False) == 5


def test_xsqlite_onecolumn_of_integer(sqlite_buildin):
    # GH 3628
    # a column_of_integers dataframe should transfer well to sql

    mono_df = DataFrame([1, 2], columns=["c0"])
    assert sql.to_sql(mono_df, con=sqlite_buildin, name="mono_df", index=False) == 2
    # computing the sum via sql
    con_x = sqlite_buildin
    the_sum = sum(my_c0[0] for my_c0 in con_x.execute("select * from mono_df"))
    # it should not fail, and gives 3 ( Issue #3628 )
    assert the_sum == 3

    result = sql.read_sql("select * from mono_df", con_x)
    tm.assert_frame_equal(result, mono_df)


def test_xsqlite_if_exists(sqlite_buildin):
    df_if_exists_1 = DataFrame({"col1": [1, 2], "col2": ["A", "B"]})
    df_if_exists_2 = DataFrame({"col1": [3, 4, 5], "col2": ["C", "D", "E"]})
    table_name = "table_if_exists"
    sql_select = f"SELECT * FROM {table_name}"

    msg = "'notvalidvalue' is not valid for if_exists"
    with pytest.raises(ValueError, match=msg):
        sql.to_sql(
            frame=df_if_exists_1,
            con=sqlite_buildin,
            name=table_name,
            if_exists="notvalidvalue",
        )
    drop_table(table_name, sqlite_buildin)

    # test if_exists='fail'
    sql.to_sql(
        frame=df_if_exists_1, con=sqlite_buildin, name=table_name, if_exists="fail"
    )
    msg = "Table 'table_if_exists' already exists"
    with pytest.raises(ValueError, match=msg):
        sql.to_sql(
            frame=df_if_exists_1,
            con=sqlite_buildin,
            name=table_name,
            if_exists="fail",
        )
    # test if_exists='replace'
    sql.to_sql(
        frame=df_if_exists_1,
        con=sqlite_buildin,
        name=table_name,
        if_exists="replace",
        index=False,
    )
    assert tquery(sql_select, con=sqlite_buildin) == [(1, "A"), (2, "B")]
    assert (
        sql.to_sql(
            frame=df_if_exists_2,
            con=sqlite_buildin,
            name=table_name,
            if_exists="replace",
            index=False,
        )
        == 3
    )
    assert tquery(sql_select, con=sqlite_buildin) == [(3, "C"), (4, "D"), (5, "E")]
    drop_table(table_name, sqlite_buildin)

    # test if_exists='append'
    assert (
        sql.to_sql(
            frame=df_if_exists_1,
            con=sqlite_buildin,
            name=table_name,
            if_exists="fail",
            index=False,
        )
        == 2
    )
    assert tquery(sql_select, con=sqlite_buildin) == [(1, "A"), (2, "B")]
    assert (
        sql.to_sql(
            frame=df_if_exists_2,
            con=sqlite_buildin,
            name=table_name,
            if_exists="append",
            index=False,
        )
        == 3
    )
    assert tquery(sql_select, con=sqlite_buildin) == [
        (1, "A"),
        (2, "B"),
        (3, "C"),
        (4, "D"),
        (5, "E"),
    ]
    drop_table(table_name, sqlite_buildin)