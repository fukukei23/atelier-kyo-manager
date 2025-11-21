
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\urllib3\contrib\securetransport.py ===
"""
SecureTranport support for urllib3 via ctypes.

This makes platform-native TLS available to urllib3 users on macOS without the
use of a compiler. This is an important feature because the Python Package
Index is moving to become a TLSv1.2-or-higher server, and the default OpenSSL
that ships with macOS is not capable of doing TLSv1.2. The only way to resolve
this is to give macOS users an alternative solution to the problem, and that
solution is to use SecureTransport.

We use ctypes here because this solution must not require a compiler. That's
because pip is not allowed to require a compiler either.

This is not intended to be a seriously long-term solution to this problem.
The hope is that PEP 543 will eventually solve this issue for us, at which
point we can retire this contrib module. But in the short term, we need to
solve the impending tire fire that is Python on Mac without this kind of
contrib module. So...here we are.

To use this module, simply import and inject it::

    import pip._vendor.urllib3.contrib.securetransport as securetransport
    securetransport.inject_into_urllib3()

Happy TLSing!

This code is a bastardised version of the code found in Will Bond's oscrypto
library. An enormous debt is owed to him for blazing this trail for us. For
that reason, this code should be considered to be covered both by urllib3's
license and by oscrypto's:

.. code-block::

    Copyright (c) 2015-2016 Will Bond <will@wbond.net>

    Permission is hereby granted, free of charge, to any person obtaining a
    copy of this software and associated documentation files (the "Software"),
    to deal in the Software without restriction, including without limitation
    the rights to use, copy, modify, merge, publish, distribute, sublicense,
    and/or sell copies of the Software, and to permit persons to whom the
    Software is furnished to do so, subject to the following conditions:

    The above copyright notice and this permission notice shall be included in
    all copies or substantial portions of the Software.

    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
    AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
    LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
    FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
    DEALINGS IN THE SOFTWARE.
"""
from __future__ import absolute_import

import contextlib
import ctypes
import errno
import os.path
import shutil
import socket
import ssl
import struct
import threading
import weakref

from .. import util
from ..packages import six
from ..util.ssl_ import PROTOCOL_TLS_CLIENT
from ._securetransport.bindings import CoreFoundation, Security, SecurityConst
from ._securetransport.low_level import (
    _assert_no_error,
    _build_tls_unknown_ca_alert,
    _cert_array_from_pem,
    _create_cfstring_array,
    _load_client_cert_chain,
    _temporary_keychain,
)

try:  # Platform-specific: Python 2
    from socket import _fileobject
except ImportError:  # Platform-specific: Python 3
    _fileobject = None
    from ..packages.backports.makefile import backport_makefile

__all__ = ["inject_into_urllib3", "extract_from_urllib3"]

# SNI always works
HAS_SNI = True

orig_util_HAS_SNI = util.HAS_SNI
orig_util_SSLContext = util.ssl_.SSLContext

# This dictionary is used by the read callback to obtain a handle to the
# calling wrapped socket. This is a pretty silly approach, but for now it'll
# do. I feel like I should be able to smuggle a handle to the wrapped socket
# directly in the SSLConnectionRef, but for now this approach will work I
# guess.
#
# We need to lock around this structure for inserts, but we don't do it for
# reads/writes in the callbacks. The reasoning here goes as follows:
#
#    1. It is not possible to call into the callbacks before the dictionary is
#       populated, so once in the callback the id must be in the dictionary.
#    2. The callbacks don't mutate the dictionary, they only read from it, and
#       so cannot conflict with any of the insertions.
#
# This is good: if we had to lock in the callbacks we'd drastically slow down
# the performance of this code.
_connection_refs = weakref.WeakValueDictionary()
_connection_ref_lock = threading.Lock()

# Limit writes to 16kB. This is OpenSSL's limit, but we'll cargo-cult it over
# for no better reason than we need *a* limit, and this one is right there.
SSL_WRITE_BLOCKSIZE = 16384

# This is our equivalent of util.ssl_.DEFAULT_CIPHERS, but expanded out to
# individual cipher suites. We need to do this because this is how
# SecureTransport wants them.
CIPHER_SUITES = [
    SecurityConst.TLS_ECDHE_ECDSA_WITH_AES_256_GCM_SHA384,
    SecurityConst.TLS_ECDHE_ECDSA_WITH_AES_128_GCM_SHA256,
    SecurityConst.TLS_ECDHE_RSA_WITH_AES_256_GCM_SHA384,
    SecurityConst.TLS_ECDHE_RSA_WITH_AES_128_GCM_SHA256,
    SecurityConst.TLS_ECDHE_ECDSA_WITH_CHACHA20_POLY1305_SHA256,
    SecurityConst.TLS_ECDHE_RSA_WITH_CHACHA20_POLY1305_SHA256,
    SecurityConst.TLS_DHE_RSA_WITH_AES_256_GCM_SHA384,
    SecurityConst.TLS_DHE_RSA_WITH_AES_128_GCM_SHA256,
    SecurityConst.TLS_ECDHE_ECDSA_WITH_AES_256_CBC_SHA384,
    SecurityConst.TLS_ECDHE_ECDSA_WITH_AES_256_CBC_SHA,
    SecurityConst.TLS_ECDHE_ECDSA_WITH_AES_128_CBC_SHA256,
    SecurityConst.TLS_ECDHE_ECDSA_WITH_AES_128_CBC_SHA,
    SecurityConst.TLS_ECDHE_RSA_WITH_AES_256_CBC_SHA384,
    SecurityConst.TLS_ECDHE_RSA_WITH_AES_256_CBC_SHA,
    SecurityConst.TLS_ECDHE_RSA_WITH_AES_128_CBC_SHA256,
    SecurityConst.TLS_ECDHE_RSA_WITH_AES_128_CBC_SHA,
    SecurityConst.TLS_DHE_RSA_WITH_AES_256_CBC_SHA256,
    SecurityConst.TLS_DHE_RSA_WITH_AES_256_CBC_SHA,
    SecurityConst.TLS_DHE_RSA_WITH_AES_128_CBC_SHA256,
    SecurityConst.TLS_DHE_RSA_WITH_AES_128_CBC_SHA,
    SecurityConst.TLS_AES_256_GCM_SHA384,
    SecurityConst.TLS_AES_128_GCM_SHA256,
    SecurityConst.TLS_RSA_WITH_AES_256_GCM_SHA384,
    SecurityConst.TLS_RSA_WITH_AES_128_GCM_SHA256,
    SecurityConst.TLS_AES_128_CCM_8_SHA256,
    SecurityConst.TLS_AES_128_CCM_SHA256,
    SecurityConst.TLS_RSA_WITH_AES_256_CBC_SHA256,
    SecurityConst.TLS_RSA_WITH_AES_128_CBC_SHA256,
    SecurityConst.TLS_RSA_WITH_AES_256_CBC_SHA,
    SecurityConst.TLS_RSA_WITH_AES_128_CBC_SHA,
]

# Basically this is simple: for PROTOCOL_SSLv23 we turn it into a low of
# TLSv1 and a high of TLSv1.2. For everything else, we pin to that version.
# TLSv1 to 1.2 are supported on macOS 10.8+
_protocol_to_min_max = {
    util.PROTOCOL_TLS: (SecurityConst.kTLSProtocol1, SecurityConst.kTLSProtocol12),
    PROTOCOL_TLS_CLIENT: (SecurityConst.kTLSProtocol1, SecurityConst.kTLSProtocol12),
}

if hasattr(ssl, "PROTOCOL_SSLv2"):
    _protocol_to_min_max[ssl.PROTOCOL_SSLv2] = (
        SecurityConst.kSSLProtocol2,
        SecurityConst.kSSLProtocol2,
    )
if hasattr(ssl, "PROTOCOL_SSLv3"):
    _protocol_to_min_max[ssl.PROTOCOL_SSLv3] = (
        SecurityConst.kSSLProtocol3,
        SecurityConst.kSSLProtocol3,
    )
if hasattr(ssl, "PROTOCOL_TLSv1"):
    _protocol_to_min_max[ssl.PROTOCOL_TLSv1] = (
        SecurityConst.kTLSProtocol1,
        SecurityConst.kTLSProtocol1,
    )
if hasattr(ssl, "PROTOCOL_TLSv1_1"):
    _protocol_to_min_max[ssl.PROTOCOL_TLSv1_1] = (
        SecurityConst.kTLSProtocol11,
        SecurityConst.kTLSProtocol11,
    )
if hasattr(ssl, "PROTOCOL_TLSv1_2"):
    _protocol_to_min_max[ssl.PROTOCOL_TLSv1_2] = (
        SecurityConst.kTLSProtocol12,
        SecurityConst.kTLSProtocol12,
    )


def inject_into_urllib3():
    """
    Monkey-patch urllib3 with SecureTransport-backed SSL-support.
    """
    util.SSLContext = SecureTransportContext
    util.ssl_.SSLContext = SecureTransportContext
    util.HAS_SNI = HAS_SNI
    util.ssl_.HAS_SNI = HAS_SNI
    util.IS_SECURETRANSPORT = True
    util.ssl_.IS_SECURETRANSPORT = True


def extract_from_urllib3():
    """
    Undo monkey-patching by :func:`inject_into_urllib3`.
    """
    util.SSLContext = orig_util_SSLContext
    util.ssl_.SSLContext = orig_util_SSLContext
    util.HAS_SNI = orig_util_HAS_SNI
    util.ssl_.HAS_SNI = orig_util_HAS_SNI
    util.IS_SECURETRANSPORT = False
    util.ssl_.IS_SECURETRANSPORT = False


def _read_callback(connection_id, data_buffer, data_length_pointer):
    """
    SecureTransport read callback. This is called by ST to request that data
    be returned from the socket.
    """
    wrapped_socket = None
    try:
        wrapped_socket = _connection_refs.get(connection_id)
        if wrapped_socket is None:
            return SecurityConst.errSSLInternal
        base_socket = wrapped_socket.socket

        requested_length = data_length_pointer[0]

        timeout = wrapped_socket.gettimeout()
        error = None
        read_count = 0

        try:
            while read_count < requested_length:
                if timeout is None or timeout >= 0:
                    if not util.wait_for_read(base_socket, timeout):
                        raise socket.error(errno.EAGAIN, "timed out")

                remaining = requested_length - read_count
                buffer = (ctypes.c_char * remaining).from_address(
                    data_buffer + read_count
                )
                chunk_size = base_socket.recv_into(buffer, remaining)
                read_count += chunk_size
                if not chunk_size:
                    if not read_count:
                        return SecurityConst.errSSLClosedGraceful
                    break
        except (socket.error) as e:
            error = e.errno

            if error is not None and error != errno.EAGAIN:
                data_length_pointer[0] = read_count
                if error == errno.ECONNRESET or error == errno.EPIPE:
                    return SecurityConst.errSSLClosedAbort
                raise

        data_length_pointer[0] = read_count

        if read_count != requested_length:
            return SecurityConst.errSSLWouldBlock

        return 0
    except Exception as e:
        if wrapped_socket is not None:
            wrapped_socket._exception = e
        return SecurityConst.errSSLInternal


def _write_callback(connection_id, data_buffer, data_length_pointer):
    """
    SecureTransport write callback. This is called by ST to request that data
    actually be sent on the network.
    """
    wrapped_socket = None
    try:
        wrapped_socket = _connection_refs.get(connection_id)
        if wrapped_socket is None:
            return SecurityConst.errSSLInternal
        base_socket = wrapped_socket.socket

        bytes_to_write = data_length_pointer[0]
        data = ctypes.string_at(data_buffer, bytes_to_write)

        timeout = wrapped_socket.gettimeout()
        error = None
        sent = 0

        try:
            while sent < bytes_to_write:
                if timeout is None or timeout >= 0:
                    if not util.wait_for_write(base_socket, timeout):
                        raise socket.error(errno.EAGAIN, "timed out")
                chunk_sent = base_socket.send(data)
                sent += chunk_sent

                # This has some needless copying here, but I'm not sure there's
                # much value in optimising this data path.
                data = data[chunk_sent:]
        except (socket.error) as e:
            error = e.errno

            if error is not None and error != errno.EAGAIN:
                data_length_pointer[0] = sent
                if error == errno.ECONNRESET or error == errno.EPIPE:
                    return SecurityConst.errSSLClosedAbort
                raise

        data_length_pointer[0] = sent

        if sent != bytes_to_write:
            return SecurityConst.errSSLWouldBlock

        return 0
    except Exception as e:
        if wrapped_socket is not None:
            wrapped_socket._exception = e
        return SecurityConst.errSSLInternal


# We need to keep these two objects references alive: if they get GC'd while
# in use then SecureTransport could attempt to call a function that is in freed
# memory. That would be...uh...bad. Yeah, that's the word. Bad.
_read_callback_pointer = Security.SSLReadFunc(_read_callback)
_write_callback_pointer = Security.SSLWriteFunc(_write_callback)


class WrappedSocket(object):
    """
    API-compatibility wrapper for Python's OpenSSL wrapped socket object.

    Note: _makefile_refs, _drop(), and _reuse() are needed for the garbage
    collector of PyPy.
    """

    def __init__(self, socket):
        self.socket = socket
        self.context = None
        self._makefile_refs = 0
        self._closed = False
        self._exception = None
        self._keychain = None
        self._keychain_dir = None
        self._client_cert_chain = None

        # We save off the previously-configured timeout and then set it to
        # zero. This is done because we use select and friends to handle the
        # timeouts, but if we leave the timeout set on the lower socket then
        # Python will "kindly" call select on that socket again for us. Avoid
        # that by forcing the timeout to zero.
        self._timeout = self.socket.gettimeout()
        self.socket.settimeout(0)

    @contextlib.contextmanager
    def _raise_on_error(self):
        """
        A context manager that can be used to wrap calls that do I/O from
        SecureTransport. If any of the I/O callbacks hit an exception, this
        context manager will correctly propagate the exception after the fact.
        This avoids silently swallowing those exceptions.

        It also correctly forces the socket closed.
        """
        self._exception = None

        # We explicitly don't catch around this yield because in the unlikely
        # event that an exception was hit in the block we don't want to swallow
        # it.
        yield
        if self._exception is not None:
            exception, self._exception = self._exception, None
            self.close()
            raise exception

    def _set_ciphers(self):
        """
        Sets up the allowed ciphers. By default this matches the set in
        util.ssl_.DEFAULT_CIPHERS, at least as supported by macOS. This is done
        custom and doesn't allow changing at this time, mostly because parsing
        OpenSSL cipher strings is going to be a freaking nightmare.
        """
        ciphers = (Security.SSLCipherSuite * len(CIPHER_SUITES))(*CIPHER_SUITES)
        result = Security.SSLSetEnabledCiphers(
            self.context, ciphers, len(CIPHER_SUITES)
        )
        _assert_no_error(result)

    def _set_alpn_protocols(self, protocols):
        """
        Sets up the ALPN protocols on the context.
        """
        if not protocols:
            return
        protocols_arr = _create_cfstring_array(protocols)
        try:
            result = Security.SSLSetALPNProtocols(self.context, protocols_arr)
            _assert_no_error(result)
        finally:
            CoreFoundation.CFRelease(protocols_arr)

    def _custom_validate(self, verify, trust_bundle):
        """
        Called when we have set custom validation. We do this in two cases:
        first, when cert validation is entirely disabled; and second, when
        using a custom trust DB.
        Raises an SSLError if the connection is not trusted.
        """
        # If we disabled cert validation, just say: cool.
        if not verify:
            return

        successes = (
            SecurityConst.kSecTrustResultUnspecified,
            SecurityConst.kSecTrustResultProceed,
        )
        try:
            trust_result = self._evaluate_trust(trust_bundle)
            if trust_result in successes:
                return
            reason = "error code: %d" % (trust_result,)
        except Exception as e:
            # Do not trust on error
            reason = "exception: %r" % (e,)

        # SecureTransport does not send an alert nor shuts down the connection.
        rec = _build_tls_unknown_ca_alert(self.version())
        self.socket.sendall(rec)
        # close the connection immediately
        # l_onoff = 1, activate linger
        # l_linger = 0, linger for 0 seoncds
        opts = struct.pack("ii", 1, 0)
        self.socket.setsockopt(socket.SOL_SOCKET, socket.SO_LINGER, opts)
        self.close()
        raise ssl.SSLError("certificate verify failed, %s" % reason)

    def _evaluate_trust(self, trust_bundle):
        # We want data in memory, so load it up.
        if os.path.isfile(trust_bundle):
            with open(trust_bundle, "rb") as f:
                trust_bundle = f.read()

        cert_array = None
        trust = Security.SecTrustRef()

        try:
            # Get a CFArray that contains the certs we want.
            cert_array = _cert_array_from_pem(trust_bundle)

            # Ok, now the hard part. We want to get the SecTrustRef that ST has
            # created for this connection, shove our CAs into it, tell ST to
            # ignore everything else it knows, and then ask if it can build a
            # chain. This is a buuuunch of code.
            result = Security.SSLCopyPeerTrust(self.context, ctypes.byref(trust))
            _assert_no_error(result)
            if not trust:
                raise ssl.SSLError("Failed to copy trust reference")

            result = Security.SecTrustSetAnchorCertificates(trust, cert_array)
            _assert_no_error(result)

            result = Security.SecTrustSetAnchorCertificatesOnly(trust, True)
            _assert_no_error(result)

            trust_result = Security.SecTrustResultType()
            result = Security.SecTrustEvaluate(trust, ctypes.byref(trust_result))
            _assert_no_error(result)
        finally:
            if trust:
                CoreFoundation.CFRelease(trust)

            if cert_array is not None:
                CoreFoundation.CFRelease(cert_array)

        return trust_result.value

    def handshake(
        self,
        server_hostname,
        verify,
        trust_bundle,
        min_version,
        max_version,
        client_cert,
        client_key,
        client_key_passphrase,
        alpn_protocols,
    ):
        """
        Actually performs the TLS handshake. This is run automatically by
        wrapped socket, and shouldn't be needed in user code.
        """
        # First, we do the initial bits of connection setup. We need to create
        # a context, set its I/O funcs, and set the connection reference.
        self.context = Security.SSLCreateContext(
            None, SecurityConst.kSSLClientSide, SecurityConst.kSSLStreamType
        )
        result = Security.SSLSetIOFuncs(
            self.context, _read_callback_pointer, _write_callback_pointer
        )
        _assert_no_error(result)

        # Here we need to compute the handle to use. We do this by taking the
        # id of self modulo 2**31 - 1. If this is already in the dictionary, we
        # just keep incrementing by one until we find a free space.
        with _connection_ref_lock:
            handle = id(self) % 2147483647
            while handle in _connection_refs:
                handle = (handle + 1) % 2147483647
            _connection_refs[handle] = self

        result = Security.SSLSetConnection(self.context, handle)
        _assert_no_error(result)

        # If we have a server hostname, we should set that too.
        if server_hostname:
            if not isinstance(server_hostname, bytes):
                server_hostname = server_hostname.encode("utf-8")

            result = Security.SSLSetPeerDomainName(
                self.context, server_hostname, len(server_hostname)
            )
            _assert_no_error(result)

        # Setup the ciphers.
        self._set_ciphers()

        # Setup the ALPN protocols.
        self._set_alpn_protocols(alpn_protocols)

        # Set the minimum and maximum TLS versions.
        result = Security.SSLSetProtocolVersionMin(self.context, min_version)
        _assert_no_error(result)

        result = Security.SSLSetProtocolVersionMax(self.context, max_version)
        _assert_no_error(result)

        # If there's a trust DB, we need to use it. We do that by telling
        # SecureTransport to break on server auth. We also do that if we don't
        # want to validate the certs at all: we just won't actually do any
        # authing in that case.
        if not verify or trust_bundle is not None:
            result = Security.SSLSetSessionOption(
                self.context, SecurityConst.kSSLSessionOptionBreakOnServerAuth, True
            )
            _assert_no_error(result)

        # If there's a client cert, we need to use it.
        if client_cert:
            self._keychain, self._keychain_dir = _temporary_keychain()
            self._client_cert_chain = _load_client_cert_chain(
                self._keychain, client_cert, client_key
            )
            result = Security.SSLSetCertificate(self.context, self._client_cert_chain)
            _assert_no_error(result)

        while True:
            with self._raise_on_error():
                result = Security.SSLHandshake(self.context)

                if result == SecurityConst.errSSLWouldBlock:
                    raise socket.timeout("handshake timed out")
                elif result == SecurityConst.errSSLServerAuthCompleted:
                    self._custom_validate(verify, trust_bundle)
                    continue
                else:
                    _assert_no_error(result)
                    break

    def fileno(self):
        return self.socket.fileno()

    # Copy-pasted from Python 3.5 source code
    def _decref_socketios(self):
        if self._makefile_refs > 0:
            self._makefile_refs -= 1
        if self._closed:
            self.close()

    def recv(self, bufsiz):
        buffer = ctypes.create_string_buffer(bufsiz)
        bytes_read = self.recv_into(buffer, bufsiz)
        data = buffer[:bytes_read]
        return data

    def recv_into(self, buffer, nbytes=None):
        # Read short on EOF.
        if self._closed:
            return 0

        if nbytes is None:
            nbytes = len(buffer)

        buffer = (ctypes.c_char * nbytes).from_buffer(buffer)
        processed_bytes = ctypes.c_size_t(0)

        with self._raise_on_error():
            result = Security.SSLRead(
                self.context, buffer, nbytes, ctypes.byref(processed_bytes)
            )

        # There are some result codes that we want to treat as "not always
        # errors". Specifically, those are errSSLWouldBlock,
        # errSSLClosedGraceful, and errSSLClosedNoNotify.
        if result == SecurityConst.errSSLWouldBlock:
            # If we didn't process any bytes, then this was just a time out.
            # However, we can get errSSLWouldBlock in situations when we *did*
            # read some data, and in those cases we should just read "short"
            # and return.
            if processed_bytes.value == 0:
                # Timed out, no data read.
                raise socket.timeout("recv timed out")
        elif result in (
            SecurityConst.errSSLClosedGraceful,
            SecurityConst.errSSLClosedNoNotify,
        ):
            # The remote peer has closed this connection. We should do so as
            # well. Note that we don't actually return here because in
            # principle this could actually be fired along with return data.
            # It's unlikely though.
            self.close()
        else:
            _assert_no_error(result)

        # Ok, we read and probably succeeded. We should return whatever data
        # was actually read.
        return processed_bytes.value

    def settimeout(self, timeout):
        self._timeout = timeout

    def gettimeout(self):
        return self._timeout

    def send(self, data):
        processed_bytes = ctypes.c_size_t(0)

        with self._raise_on_error():
            result = Security.SSLWrite(
                self.context, data, len(data), ctypes.byref(processed_bytes)
            )

        if result == SecurityConst.errSSLWouldBlock and processed_bytes.value == 0:
            # Timed out
            raise socket.timeout("send timed out")
        else:
            _assert_no_error(result)

        # We sent, and probably succeeded. Tell them how much we sent.
        return processed_bytes.value

    def sendall(self, data):
        total_sent = 0
        while total_sent < len(data):
            sent = self.send(data[total_sent : total_sent + SSL_WRITE_BLOCKSIZE])
            total_sent += sent

    def shutdown(self):
        with self._raise_on_error():
            Security.SSLClose(self.context)

    def close(self):
        # TODO: should I do clean shutdown here? Do I have to?
        if self._makefile_refs < 1:
            self._closed = True
            if self.context:
                CoreFoundation.CFRelease(self.context)
                self.context = None
            if self._client_cert_chain:
                CoreFoundation.CFRelease(self._client_cert_chain)
                self._client_cert_chain = None
            if self._keychain:
                Security.SecKeychainDelete(self._keychain)
                CoreFoundation.CFRelease(self._keychain)
                shutil.rmtree(self._keychain_dir)
                self._keychain = self._keychain_dir = None
            return self.socket.close()
        else:
            self._makefile_refs -= 1

    def getpeercert(self, binary_form=False):
        # Urgh, annoying.
        #
        # Here's how we do this:
        #
        # 1. Call SSLCopyPeerTrust to get hold of the trust object for this
        #    connection.
        # 2. Call SecTrustGetCertificateAtIndex for index 0 to get the leaf.
        # 3. To get the CN, call SecCertificateCopyCommonName and process that
        #    string so that it's of the appropriate type.
        # 4. To get the SAN, we need to do something a bit more complex:
        #    a. Call SecCertificateCopyValues to get the data, requesting
        #       kSecOIDSubjectAltName.
        #    b. Mess about with this dictionary to try to get the SANs out.
        #
        # This is gross. Really gross. It's going to be a few hundred LoC extra
        # just to repeat something that SecureTransport can *already do*. So my
        # operating assumption at this time is that what we want to do is
        # instead to just flag to urllib3 that it shouldn't do its own hostname
        # validation when using SecureTransport.
        if not binary_form:
            raise ValueError("SecureTransport only supports dumping binary certs")
        trust = Security.SecTrustRef()
        certdata = None
        der_bytes = None

        try:
            # Grab the trust store.
            result = Security.SSLCopyPeerTrust(self.context, ctypes.byref(trust))
            _assert_no_error(result)
            if not trust:
                # Probably we haven't done the handshake yet. No biggie.
                return None

            cert_count = Security.SecTrustGetCertificateCount(trust)
            if not cert_count:
                # Also a case that might happen if we haven't handshaked.
                # Handshook? Handshaken?
                return None

            leaf = Security.SecTrustGetCertificateAtIndex(trust, 0)
            assert leaf

            # Ok, now we want the DER bytes.
            certdata = Security.SecCertificateCopyData(leaf)
            assert certdata

            data_length = CoreFoundation.CFDataGetLength(certdata)
            data_buffer = CoreFoundation.CFDataGetBytePtr(certdata)
            der_bytes = ctypes.string_at(data_buffer, data_length)
        finally:
            if certdata:
                CoreFoundation.CFRelease(certdata)
            if trust:
                CoreFoundation.CFRelease(trust)

        return der_bytes

    def version(self):
        protocol = Security.SSLProtocol()
        result = Security.SSLGetNegotiatedProtocolVersion(
            self.context, ctypes.byref(protocol)
        )
        _assert_no_error(result)
        if protocol.value == SecurityConst.kTLSProtocol13:
            raise ssl.SSLError("SecureTransport does not support TLS 1.3")
        elif protocol.value == SecurityConst.kTLSProtocol12:
            return "TLSv1.2"
        elif protocol.value == SecurityConst.kTLSProtocol11:
            return "TLSv1.1"
        elif protocol.value == SecurityConst.kTLSProtocol1:
            return "TLSv1"
        elif protocol.value == SecurityConst.kSSLProtocol3:
            return "SSLv3"
        elif protocol.value == SecurityConst.kSSLProtocol2:
            return "SSLv2"
        else:
            raise ssl.SSLError("Unknown TLS version: %r" % protocol)

    def _reuse(self):
        self._makefile_refs += 1

    def _drop(self):
        if self._makefile_refs < 1:
            self.close()
        else:
            self._makefile_refs -= 1


if _fileobject:  # Platform-specific: Python 2

    def makefile(self, mode, bufsize=-1):
        self._makefile_refs += 1
        return _fileobject(self, mode, bufsize, close=True)

else:  # Platform-specific: Python 3

    def makefile(self, mode="r", buffering=None, *args, **kwargs):
        # We disable buffering with SecureTransport because it conflicts with
        # the buffering that ST does internally (see issue #1153 for more).
        buffering = 0
        return backport_makefile(self, mode, buffering, *args, **kwargs)


WrappedSocket.makefile = makefile


class SecureTransportContext(object):
    """
    I am a wrapper class for the SecureTransport library, to translate the
    interface of the standard library ``SSLContext`` object to calls into
    SecureTransport.
    """

    def __init__(self, protocol):
        self._min_version, self._max_version = _protocol_to_min_max[protocol]
        self._options = 0
        self._verify = False
        self._trust_bundle = None
        self._client_cert = None
        self._client_key = None
        self._client_key_passphrase = None
        self._alpn_protocols = None

    @property
    def check_hostname(self):
        """
        SecureTransport cannot have its hostname checking disabled. For more,
        see the comment on getpeercert() in this file.
        """
        return True

    @check_hostname.setter
    def check_hostname(self, value):
        """
        SecureTransport cannot have its hostname checking disabled. For more,
        see the comment on getpeercert() in this file.
        """
        pass

    @property
    def options(self):
        # TODO: Well, crap.
        #
        # So this is the bit of the code that is the most likely to cause us
        # trouble. Essentially we need to enumerate all of the SSL options that
        # users might want to use and try to see if we can sensibly translate
        # them, or whether we should just ignore them.
        return self._options

    @options.setter
    def options(self, value):
        # TODO: Update in line with above.
        self._options = value

    @property
    def verify_mode(self):
        return ssl.CERT_REQUIRED if self._verify else ssl.CERT_NONE

    @verify_mode.setter
    def verify_mode(self, value):
        self._verify = True if value == ssl.CERT_REQUIRED else False

    def set_default_verify_paths(self):
        # So, this has to do something a bit weird. Specifically, what it does
        # is nothing.
        #
        # This means that, if we had previously had load_verify_locations
        # called, this does not undo that. We need to do that because it turns
        # out that the rest of the urllib3 code will attempt to load the
        # default verify paths if it hasn't been told about any paths, even if
        # the context itself was sometime earlier. We resolve that by just
        # ignoring it.
        pass

    def load_default_certs(self):
        return self.set_default_verify_paths()

    def set_ciphers(self, ciphers):
        # For now, we just require the default cipher string.
        if ciphers != util.ssl_.DEFAULT_CIPHERS:
            raise ValueError("SecureTransport doesn't support custom cipher strings")

    def load_verify_locations(self, cafile=None, capath=None, cadata=None):
        # OK, we only really support cadata and cafile.
        if capath is not None:
            raise ValueError("SecureTransport does not support cert directories")

        # Raise if cafile does not exist.
        if cafile is not None:
            with open(cafile):
                pass

        self._trust_bundle = cafile or cadata

    def load_cert_chain(self, certfile, keyfile=None, password=None):
        self._client_cert = certfile
        self._client_key = keyfile
        self._client_cert_passphrase = password

    def set_alpn_protocols(self, protocols):
        """
        Sets the ALPN protocols that will later be set on the context.

        Raises a NotImplementedError if ALPN is not supported.
        """
        if not hasattr(Security, "SSLSetALPNProtocols"):
            raise NotImplementedError(
                "SecureTransport supports ALPN only in macOS 10.12+"
            )
        self._alpn_protocols = [six.ensure_binary(p) for p in protocols]

    def wrap_socket(
        self,
        sock,
        server_side=False,
        do_handshake_on_connect=True,
        suppress_ragged_eofs=True,
        server_hostname=None,
    ):
        # So, what do we do here? Firstly, we assert some properties. This is a
        # stripped down shim, so there is some functionality we don't support.
        # See PEP 543 for the real deal.
        assert not server_side
        assert do_handshake_on_connect
        assert suppress_ragged_eofs

        # Ok, we're good to go. Now we want to create the wrapped socket object
        # and store it in the appropriate place.
        wrapped_socket = WrappedSocket(sock)

        # Now we can handshake
        wrapped_socket.handshake(
            server_hostname,
            self._verify,
            self._trust_bundle,
            self._min_version,
            self._max_version,
            self._client_cert,
            self._client_key,
            self._client_key_passphrase,
            self._alpn_protocols,
        )
        return wrapped_socket

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\et_xmlfile\incremental_tree.py ===
# Code modified from cPython's Lib/xml/etree/ElementTree.py
# The write() code is modified to allow specifying a particular namespace
# uri -> prefix mapping.
#
# ---------------------------------------------------------------------
# Licensed to PSF under a Contributor Agreement.
# See https://www.python.org/psf/license for licensing details.
#
# ElementTree
# Copyright (c) 1999-2008 by Fredrik Lundh.  All rights reserved.
#
# fredrik@pythonware.com
# http://www.pythonware.com
# --------------------------------------------------------------------
# The ElementTree toolkit is
#
# Copyright (c) 1999-2008 by Fredrik Lundh
#
# By obtaining, using, and/or copying this software and/or its
# associated documentation, you agree that you have read, understood,
# and will comply with the following terms and conditions:
#
# Permission to use, copy, modify, and distribute this software and
# its associated documentation for any purpose and without fee is
# hereby granted, provided that the above copyright notice appears in
# all copies, and that both that copyright notice and this permission
# notice appear in supporting documentation, and that the name of
# Secret Labs AB or the author not be used in advertising or publicity
# pertaining to distribution of the software without specific, written
# prior permission.
#
# SECRET LABS AB AND THE AUTHOR DISCLAIMS ALL WARRANTIES WITH REGARD
# TO THIS SOFTWARE, INCLUDING ALL IMPLIED WARRANTIES OF MERCHANT-
# ABILITY AND FITNESS.  IN NO EVENT SHALL SECRET LABS AB OR THE AUTHOR
# BE LIABLE FOR ANY SPECIAL, INDIRECT OR CONSEQUENTIAL DAMAGES OR ANY
# DAMAGES WHATSOEVER RESULTING FROM LOSS OF USE, DATA OR PROFITS,
# WHETHER IN AN ACTION OF CONTRACT, NEGLIGENCE OR OTHER TORTIOUS
# ACTION, ARISING OUT OF OR IN CONNECTION WITH THE USE OR PERFORMANCE
# OF THIS SOFTWARE.
# --------------------------------------------------------------------
import contextlib
import io

import xml.etree.ElementTree as ET


def current_global_nsmap():
    return {
        prefix: uri for uri, prefix in ET._namespace_map.items()
    }


class IncrementalTree(ET.ElementTree):

    def write(
        self,
        file_or_filename,
        encoding=None,
        xml_declaration=None,
        default_namespace=None,
        method=None,
        *,
        short_empty_elements=True,
        nsmap=None,
        root_ns_only=False,
        minimal_ns_only=False,
    ):
        """Write element tree to a file as XML.

        Arguments:
          *file_or_filename* -- file name or a file object opened for writing

          *encoding* -- the output encoding (default: US-ASCII)

          *xml_declaration* -- bool indicating if an XML declaration should be
                               added to the output. If None, an XML declaration
                               is added if encoding IS NOT either of:
                               US-ASCII, UTF-8, or Unicode

          *default_namespace* -- sets the default XML namespace (for "xmlns").
                                 Takes precedence over any default namespace
                                 provided in nsmap or
                                 xml.etree.ElementTree.register_namespace().

          *method* -- either "xml" (default), "html, "text", or "c14n"

          *short_empty_elements* -- controls the formatting of elements
                                    that contain no content. If True (default)
                                    they are emitted as a single self-closed
                                    tag, otherwise they are emitted as a pair
                                    of start/end tags

          *nsmap* -- a mapping of namespace prefixes to URIs. These take
                     precedence over any mappings registered using
                     xml.etree.ElementTree.register_namespace(). The
                     default_namespace argument, if supplied, takes precedence
                     over any default namespace supplied in nsmap. All supplied
                     namespaces will be declared on the root element, even if
                     unused in the document.

          *root_ns_only* -- bool indicating namespace declrations should only
                            be written on the root element.  This requires two
                            passes of the xml tree adding additional time to
                            the writing process. This is primarily meant to
                            mimic xml.etree.ElementTree's behaviour.

          *minimal_ns_only* -- bool indicating only namespaces that were used
                               to qualify elements or attributes should be
                               declared. All namespace declarations will be
                               written on the root element regardless of the
                               value of the root_ns_only arg. Requires two
                               passes of the xml tree adding additional time to
                               the writing process.

        """
        if not method:
            method = "xml"
        elif method not in ("text", "xml", "html"):
            raise ValueError("unknown method %r" % method)
        if not encoding:
            encoding = "us-ascii"

        with _get_writer(file_or_filename, encoding) as (write, declared_encoding):
            if method == "xml" and (
                xml_declaration
                or (
                    xml_declaration is None
                    and encoding.lower() != "unicode"
                    and declared_encoding.lower() not in ("utf-8", "us-ascii")
                )
            ):
                write("<?xml version='1.0' encoding='%s'?>\n" % (declared_encoding,))
            if method == "text":
                ET._serialize_text(write, self._root)
            else:
                if method == "xml":
                    is_html = False
                else:
                    is_html = True
                if nsmap:
                    if None in nsmap:
                        raise ValueError(
                            'Found None as default nsmap prefix in nsmap. '
                            'Use "" as the default namespace prefix.'
                        )
                    new_nsmap = nsmap.copy()
                else:
                    new_nsmap = {}
                if default_namespace:
                    new_nsmap[""] = default_namespace
                if root_ns_only or minimal_ns_only:
                    # _namespaces returns a mapping of only the namespaces that
                    # were used.
                    new_nsmap = _namespaces(
                        self._root,
                        default_namespace,
                        new_nsmap,
                    )
                    if not minimal_ns_only:
                        if nsmap:
                            # We want all namespaces defined in the provided
                            # nsmap to be declared regardless of whether
                            # they've been used.
                            new_nsmap.update(nsmap)
                        if default_namespace:
                            new_nsmap[""] = default_namespace
                global_nsmap = {
                    prefix: uri for uri, prefix in ET._namespace_map.items()
                }
                if None in global_nsmap:
                    raise ValueError(
                        'Found None as default nsmap prefix in nsmap registered with '
                        'register_namespace. Use "" for the default namespace prefix.'
                    )
                nsmap_scope = {}
                _serialize_ns_xml(
                    write,
                    self._root,
                    nsmap_scope,
                    global_nsmap,
                    is_html=is_html,
                    is_root=True,
                    short_empty_elements=short_empty_elements,
                    new_nsmap=new_nsmap,
                )


def _make_new_ns_prefix(
    nsmap_scope,
    global_prefixes,
    local_nsmap=None,
    default_namespace=None,
):
    i = len(nsmap_scope)
    if default_namespace is not None and "" not in nsmap_scope:
        # Keep the same numbering scheme as python which assumes the default
        # namespace is present if supplied.
        i += 1

    while True:
        prefix = f"ns{i}"
        if (
            prefix not in nsmap_scope
            and prefix not in global_prefixes
            and (
                not local_nsmap or prefix not in local_nsmap
            )
        ):
            return prefix
        i += 1


def _get_or_create_prefix(
    uri,
    nsmap_scope,
    global_nsmap,
    new_namespace_prefixes,
    uri_to_prefix,
    for_default_namespace_attr_prefix=False,
):
    """Find a prefix that doesn't conflict with the ns scope or create a new prefix

    This function mutates nsmap_scope, global_nsmap, new_namespace_prefixes and
    uri_to_prefix. It is intended to keep state in _serialize_ns_xml consistent
    while deduplicating the house keeping code or updating these dictionaries.
    """
    # Check if we can reuse an existing (global) prefix within the current
    # namespace scope. There maybe many prefixes pointing to a single URI by
    # this point and we need to select a prefix that is not in use in the
    # current scope.
    for global_prefix, global_uri in global_nsmap.items():
        if uri == global_uri and global_prefix not in nsmap_scope:
            prefix = global_prefix
            break
    else:  # no break
        # We couldn't find a suitable existing prefix for this namespace scope,
        # let's create a new one.
        prefix = _make_new_ns_prefix(nsmap_scope, global_prefixes=global_nsmap)
        global_nsmap[prefix] = uri
    nsmap_scope[prefix] = uri
    if not for_default_namespace_attr_prefix:
        # Don't override the actual default namespace prefix
        uri_to_prefix[uri] = prefix
    if prefix != "xml":
        new_namespace_prefixes.add(prefix)
    return prefix


def _find_default_namespace_attr_prefix(
    default_namespace,
    nsmap,
    local_nsmap,
    global_prefixes,
    provided_default_namespace=None,
):
    # Search the provided nsmap for any prefixes for this uri that aren't the
    # default namespace ""
    for prefix, uri in nsmap.items():
        if uri == default_namespace and prefix != "":
            return prefix

    for prefix, uri in local_nsmap.items():
        if uri == default_namespace and prefix != "":
            return prefix

    # _namespace_map is a 1:1 mapping of uri -> prefix
    prefix = ET._namespace_map.get(default_namespace)
    if prefix and prefix not in nsmap:
        return prefix

    return _make_new_ns_prefix(
        nsmap,
        global_prefixes,
        local_nsmap,
        provided_default_namespace,
    )


def process_attribs(
    elem,
    is_nsmap_scope_changed,
    default_ns_attr_prefix,
    nsmap_scope,
    global_nsmap,
    new_namespace_prefixes,
    uri_to_prefix,
):
    item_parts = []
    for k, v in elem.items():
        if isinstance(k, ET.QName):
            k = k.text
        try:
            if k[:1] == "{":
                uri_and_name = k[1:].rsplit("}", 1)
                try:
                    prefix = uri_to_prefix[uri_and_name[0]]
                except KeyError:
                    if not is_nsmap_scope_changed:
                        # We're about to mutate the these dicts so
                        # let's copy them first. We don't have to
                        # recompute other mappings as we're looking up
                        # or creating a new prefix
                        nsmap_scope = nsmap_scope.copy()
                        uri_to_prefix = uri_to_prefix.copy()
                        is_nsmap_scope_changed = True
                    prefix = _get_or_create_prefix(
                        uri_and_name[0],
                        nsmap_scope,
                        global_nsmap,
                        new_namespace_prefixes,
                        uri_to_prefix,
                    )

                if not prefix:
                    if default_ns_attr_prefix:
                        prefix = default_ns_attr_prefix
                    else:
                        for prefix, known_uri in nsmap_scope.items():
                            if known_uri == uri_and_name[0] and prefix != "":
                                default_ns_attr_prefix = prefix
                                break
                        else:  # no break
                            if not is_nsmap_scope_changed:
                                # We're about to mutate the these dicts so
                                # let's copy them first. We don't have to
                                # recompute other mappings as we're looking up
                                # or creating a new prefix
                                nsmap_scope = nsmap_scope.copy()
                                uri_to_prefix = uri_to_prefix.copy()
                                is_nsmap_scope_changed = True
                            prefix = _get_or_create_prefix(
                                uri_and_name[0],
                                nsmap_scope,
                                global_nsmap,
                                new_namespace_prefixes,
                                uri_to_prefix,
                                for_default_namespace_attr_prefix=True,
                            )
                            default_ns_attr_prefix = prefix
                k = f"{prefix}:{uri_and_name[1]}"
        except TypeError:
            ET._raise_serialization_error(k)

        if isinstance(v, ET.QName):
            if v.text[:1] != "{":
                v = v.text
            else:
                uri_and_name = v.text[1:].rsplit("}", 1)
                try:
                    prefix = uri_to_prefix[uri_and_name[0]]
                except KeyError:
                    if not is_nsmap_scope_changed:
                        # We're about to mutate the these dicts so
                        # let's copy them first. We don't have to
                        # recompute other mappings as we're looking up
                        # or creating a new prefix
                        nsmap_scope = nsmap_scope.copy()
                        uri_to_prefix = uri_to_prefix.copy()
                        is_nsmap_scope_changed = True
                    prefix = _get_or_create_prefix(
                        uri_and_name[0],
                        nsmap_scope,
                        global_nsmap,
                        new_namespace_prefixes,
                        uri_to_prefix,
                    )
                v = f"{prefix}:{uri_and_name[1]}"
        item_parts.append((k, v))
    return item_parts, default_ns_attr_prefix, nsmap_scope


def write_elem_start(
    write,
    elem,
    nsmap_scope,
    global_nsmap,
    short_empty_elements,
    is_html,
    is_root=False,
    uri_to_prefix=None,
    default_ns_attr_prefix=None,
    new_nsmap=None,
    **kwargs,
):
    """Write the opening tag (including self closing) and element text.

    Refer to _serialize_ns_xml for description of arguments.

    nsmap_scope should be an empty dictionary on first call. All nsmap prefixes
    must be strings with the default namespace prefix represented by "".

    eg.
    - <foo attr1="one">      (returns tag = 'foo')
    - <foo attr1="one">text  (returns tag = 'foo')
    - <foo attr1="one" />    (returns tag = None)

    Returns:
        tag:
            The tag name to be closed or None if no closing required.
        nsmap_scope:
            The current nsmap after any prefix to uri additions from this
            element. This is the input dict if unmodified or an updated copy.
        default_ns_attr_prefix:
            The prefix for the default namespace to use with attrs.
        uri_to_prefix:
            The current uri to prefix map after any uri to prefix additions
            from this element. This is the input dict if unmodified or an
            updated copy.
        next_remains_root:
            A bool indicating if the child element(s) should be treated as
            their own roots.
    """
    tag = elem.tag
    text = elem.text

    if tag is ET.Comment:
        write("<!--%s-->" % text)
        tag = None
        next_remains_root = False
    elif tag is ET.ProcessingInstruction:
        write("<?%s?>" % text)
        tag = None
        next_remains_root = False
    else:
        if new_nsmap:
            is_nsmap_scope_changed = True
            nsmap_scope = nsmap_scope.copy()
            nsmap_scope.update(new_nsmap)
            new_namespace_prefixes = set(new_nsmap.keys())
            new_namespace_prefixes.discard("xml")
            # We need to recompute the uri to prefixes
            uri_to_prefix = None
            default_ns_attr_prefix = None
        else:
            is_nsmap_scope_changed = False
            new_namespace_prefixes = set()

        if uri_to_prefix is None:
            if None in nsmap_scope:
                raise ValueError(
                    'Found None as a namespace prefix. Use "" as the default namespace prefix.'
                )
            uri_to_prefix = {uri: prefix for prefix, uri in nsmap_scope.items()}
            if "" in nsmap_scope:
                # There may be multiple prefixes for the default namespace but
                # we want to make sure we preferentially use "" (for elements)
                uri_to_prefix[nsmap_scope[""]] = ""

        if tag is None:
            # tag supression where tag is set to None
            # Don't change is_root so namespaces can be passed down
            next_remains_root = is_root
            if text:
                write(ET._escape_cdata(text))
        else:
            next_remains_root = False
            if isinstance(tag, ET.QName):
                tag = tag.text
            try:
                # These splits / fully qualified tag creationg are the
                # bottleneck in this implementation vs the python
                # implementation.
                # The following split takes ~42ns with no uri and ~85ns if a
                # prefix is present. If the uri was present, we then need to
                # look up a prefix (~14ns) and create the fully qualified
                # string (~41ns).  This gives a total of ~140ns where a uri is
                # present.
                # Python's implementation needs to preprocess the tree to
                # create a dict of qname -> tag by traversing the tree which
                # takes a bit of extra time but it quickly makes that back by
                # only having to do a dictionary look up (~14ns) for each tag /
                # attrname vs our splitting (~140ns).
                # So here we have the flexibility of being able to redefine the
                # uri a prefix points to midway through serialisation at the
                # expense of performance (~10% slower for a 1mb file on my
                # machine).
                if tag[:1] == "{":
                    uri_and_name = tag[1:].rsplit("}", 1)
                    try:
                        prefix = uri_to_prefix[uri_and_name[0]]
                    except KeyError:
                        if not is_nsmap_scope_changed:
                            # We're about to mutate the these dicts so let's
                            # copy them first. We don't have to recompute other
                            # mappings as we're looking up or creating a new
                            # prefix
                            nsmap_scope = nsmap_scope.copy()
                            uri_to_prefix = uri_to_prefix.copy()
                            is_nsmap_scope_changed = True
                        prefix = _get_or_create_prefix(
                            uri_and_name[0],
                            nsmap_scope,
                            global_nsmap,
                            new_namespace_prefixes,
                            uri_to_prefix,
                        )
                    if prefix:
                        tag = f"{prefix}:{uri_and_name[1]}"
                    else:
                        tag = uri_and_name[1]
                elif "" in nsmap_scope:
                    raise ValueError(
                        "cannot use non-qualified names with default_namespace option"
                    )
            except TypeError:
                ET._raise_serialization_error(tag)

            write("<" + tag)

            if elem.attrib:
                item_parts, default_ns_attr_prefix, nsmap_scope = process_attribs(
                    elem,
                    is_nsmap_scope_changed,
                    default_ns_attr_prefix,
                    nsmap_scope,
                    global_nsmap,
                    new_namespace_prefixes,
                    uri_to_prefix,
                )
            else:
                item_parts = []
            if new_namespace_prefixes:
                ns_attrs = []
                for k in sorted(new_namespace_prefixes):
                    v = nsmap_scope[k]
                    if k:
                        k = "xmlns:" + k
                    else:
                        k = "xmlns"
                    ns_attrs.append((k, v))
                if is_html:
                    write("".join([f' {k}="{ET._escape_attrib_html(v)}"' for k, v in ns_attrs]))
                else:
                    write("".join([f' {k}="{ET._escape_attrib(v)}"' for k, v in ns_attrs]))
            if item_parts:
                if is_html:
                    write("".join([f' {k}="{ET._escape_attrib_html(v)}"' for k, v in item_parts]))
                else:
                    write("".join([f' {k}="{ET._escape_attrib(v)}"' for k, v in item_parts]))
            if is_html:
                write(">")
                ltag = tag.lower()
                if text:
                    if ltag == "script" or ltag == "style":
                        write(text)
                    else:
                        write(ET._escape_cdata(text))
                if ltag in ET.HTML_EMPTY:
                    tag = None
            elif text or len(elem) or not short_empty_elements:
                write(">")
                if text:
                    write(ET._escape_cdata(text))
            else:
                tag = None
                write(" />")
    return (
        tag,
        nsmap_scope,
        default_ns_attr_prefix,
        uri_to_prefix,
        next_remains_root,
    )


def _serialize_ns_xml(
    write,
    elem,
    nsmap_scope,
    global_nsmap,
    short_empty_elements,
    is_html,
    is_root=False,
    uri_to_prefix=None,
    default_ns_attr_prefix=None,
    new_nsmap=None,
    **kwargs,
):
    """Serialize an element or tree using 'write' for output.

    Args:
        write:
            A function to write the xml to its destination.
        elem:
            The element to serialize.
        nsmap_scope:
            The current prefix to uri mapping for this element. This should be
            an empty dictionary for the root element. Additional namespaces are
            progressively added using the new_nsmap arg.
        global_nsmap:
            A dict copy of the globally registered _namespace_map in uri to
            prefix form
        short_empty_elements:
          Controls the formatting of elements that contain no content. If True
          (default) they are emitted as a single self-closed tag, otherwise
          they are emitted as a pair of start/end tags.
        is_html:
            Set to True to serialize as HTML otherwise XML.
        is_root:
            Boolean indicating if this is a root element.
        uri_to_prefix:
            Current state of the mapping of uri to prefix.
        default_ns_attr_prefix:
        new_nsmap:
            New prefix -> uri mapping to be applied to this element.
    """
    (
        tag,
        nsmap_scope,
        default_ns_attr_prefix,
        uri_to_prefix,
        next_remains_root,
    ) = write_elem_start(
        write,
        elem,
        nsmap_scope,
        global_nsmap,
        short_empty_elements,
        is_html,
        is_root,
        uri_to_prefix,
        default_ns_attr_prefix,
        new_nsmap=new_nsmap,
    )
    for e in elem:
        _serialize_ns_xml(
            write,
            e,
            nsmap_scope,
            global_nsmap,
            short_empty_elements,
            is_html,
            next_remains_root,
            uri_to_prefix,
            default_ns_attr_prefix,
            new_nsmap=None,
        )
    if tag:
        write(f"</{tag}>")
    if elem.tail:
        write(ET._escape_cdata(elem.tail))


def _qnames_iter(elem):
    """Iterate through all the qualified names in elem"""
    seen_el_qnames = set()
    seen_other_qnames = set()
    for this_elem in elem.iter():
        tag = this_elem.tag
        if isinstance(tag, str):
            if tag not in seen_el_qnames:
                seen_el_qnames.add(tag)
                yield tag, True
        elif isinstance(tag, ET.QName):
            tag = tag.text
            if tag not in seen_el_qnames:
                seen_el_qnames.add(tag)
                yield tag, True
        elif (
            tag is not None
            and tag is not ET.ProcessingInstruction
            and tag is not ET.Comment
        ):
            ET._raise_serialization_error(tag)

        for key, value in this_elem.items():
            if isinstance(key, ET.QName):
                key = key.text
            if key not in seen_other_qnames:
                seen_other_qnames.add(key)
                yield key, False

            if isinstance(value, ET.QName):
                if value.text not in seen_other_qnames:
                    seen_other_qnames.add(value.text)
                    yield value.text, False

        text = this_elem.text
        if isinstance(text, ET.QName):
            if text.text not in seen_other_qnames:
                seen_other_qnames.add(text.text)
                yield text.text, False


def _namespaces(
    elem,
    default_namespace=None,
    nsmap=None,
):
    """Find all namespaces used in the document and return a prefix to uri map"""
    if nsmap is None:
        nsmap = {}

    out_nsmap = {}

    seen_uri_to_prefix = {}
    # Multiple prefixes may be present for a single uri. This will select the
    # last prefix found in nsmap for a given uri.
    local_prefix_map = {uri: prefix for prefix, uri in nsmap.items()}
    if default_namespace is not None:
        local_prefix_map[default_namespace] = ""
    elif "" in nsmap:
        # but we make sure the default prefix always take precedence
        local_prefix_map[nsmap[""]] = ""

    global_prefixes = set(ET._namespace_map.values())
    has_unqual_el = False
    default_namespace_attr_prefix = None
    for qname, is_el in _qnames_iter(elem):
        try:
            if qname[:1] == "{":
                uri_and_name = qname[1:].rsplit("}", 1)

                prefix = seen_uri_to_prefix.get(uri_and_name[0])
                if prefix is None:
                    prefix = local_prefix_map.get(uri_and_name[0])
                    if prefix is None or prefix in out_nsmap:
                        prefix = ET._namespace_map.get(uri_and_name[0])
                        if prefix is None or prefix in out_nsmap:
                            prefix = _make_new_ns_prefix(
                                out_nsmap,
                                global_prefixes,
                                nsmap,
                                default_namespace,
                            )
                    if prefix or is_el:
                        out_nsmap[prefix] = uri_and_name[0]
                        seen_uri_to_prefix[uri_and_name[0]] = prefix

                if not is_el and not prefix and not default_namespace_attr_prefix:
                    # Find the alternative prefix to use with non-element
                    # names
                    default_namespace_attr_prefix = _find_default_namespace_attr_prefix(
                        uri_and_name[0],
                        out_nsmap,
                        nsmap,
                        global_prefixes,
                        default_namespace,
                    )
                    out_nsmap[default_namespace_attr_prefix] = uri_and_name[0]
                    # Don't add this uri to prefix mapping as it might override
                    # the uri -> "" default mapping. We'll fix this up at the
                    # end of the fn.
                    # local_prefix_map[uri_and_name[0]] = default_namespace_attr_prefix
            else:
                if is_el:
                    has_unqual_el = True
        except TypeError:
            ET._raise_serialization_error(qname)

    if "" in out_nsmap and has_unqual_el:
        # FIXME: can this be handled in XML 1.0?
        raise ValueError(
            "cannot use non-qualified names with default_namespace option"
        )

    # The xml prefix doesn't need to be declared but may have been used to
    # prefix names. Let's remove it if it has been used
    out_nsmap.pop("xml", None)
    return out_nsmap


def tostring(
    element,
    encoding=None,
    method=None,
    *,
    xml_declaration=None,
    default_namespace=None,
    short_empty_elements=True,
    nsmap=None,
    root_ns_only=False,
    minimal_ns_only=False,
    tree_cls=IncrementalTree,
):
    """Generate string representation of XML element.

    All subelements are included.  If encoding is "unicode", a string
    is returned. Otherwise a bytestring is returned.

    *element* is an Element instance, *encoding* is an optional output
    encoding defaulting to US-ASCII, *method* is an optional output which can
    be one of "xml" (default), "html", "text" or "c14n", *default_namespace*
    sets the default XML namespace (for "xmlns").

    Returns an (optionally) encoded string containing the XML data.

    """
    stream = io.StringIO() if encoding == "unicode" else io.BytesIO()
    tree_cls(element).write(
        stream,
        encoding,
        xml_declaration=xml_declaration,
        default_namespace=default_namespace,
        method=method,
        short_empty_elements=short_empty_elements,
        nsmap=nsmap,
        root_ns_only=root_ns_only,
        minimal_ns_only=minimal_ns_only,
    )
    return stream.getvalue()


def tostringlist(
    element,
    encoding=None,
    method=None,
    *,
    xml_declaration=None,
    default_namespace=None,
    short_empty_elements=True,
    nsmap=None,
    root_ns_only=False,
    minimal_ns_only=False,
    tree_cls=IncrementalTree,
):
    lst = []
    stream = ET._ListDataStream(lst)
    tree_cls(element).write(
        stream,
        encoding,
        xml_declaration=xml_declaration,
        default_namespace=default_namespace,
        method=method,
        short_empty_elements=short_empty_elements,
        nsmap=nsmap,
        root_ns_only=root_ns_only,
        minimal_ns_only=minimal_ns_only,
    )
    return lst


def compat_tostring(
    element,
    encoding=None,
    method=None,
    *,
    xml_declaration=None,
    default_namespace=None,
    short_empty_elements=True,
    nsmap=None,
    root_ns_only=True,
    minimal_ns_only=False,
    tree_cls=IncrementalTree,
):
    """tostring with options that produce the same results as xml.etree.ElementTree.tostring

    root_ns_only=True is a bit slower than False as it needs to traverse the
    tree one more time to collect all the namespaces.
    """
    return tostring(
        element,
        encoding=encoding,
        method=method,
        xml_declaration=xml_declaration,
        default_namespace=default_namespace,
        short_empty_elements=short_empty_elements,
        nsmap=nsmap,
        root_ns_only=root_ns_only,
        minimal_ns_only=minimal_ns_only,
        tree_cls=tree_cls,
    )


# --------------------------------------------------------------------
# serialization support

@contextlib.contextmanager
def _get_writer(file_or_filename, encoding):
    # Copied from Python 3.12
    # returns text write method and release all resources after using
    try:
        write = file_or_filename.write
    except AttributeError:
        # file_or_filename is a file name
        if encoding.lower() == "unicode":
            encoding = "utf-8"
        with open(file_or_filename, "w", encoding=encoding,
                  errors="xmlcharrefreplace") as file:
            yield file.write, encoding
    else:
        # file_or_filename is a file-like object
        # encoding determines if it is a text or binary writer
        if encoding.lower() == "unicode":
            # use a text writer as is
            yield write, getattr(file_or_filename, "encoding", None) or "utf-8"
        else:
            # wrap a binary writer with TextIOWrapper
            with contextlib.ExitStack() as stack:
                if isinstance(file_or_filename, io.BufferedIOBase):
                    file = file_or_filename
                elif isinstance(file_or_filename, io.RawIOBase):
                    file = io.BufferedWriter(file_or_filename)
                    # Keep the original file open when the BufferedWriter is
                    # destroyed
                    stack.callback(file.detach)
                else:
                    # This is to handle passed objects that aren't in the
                    # IOBase hierarchy, but just have a write method
                    file = io.BufferedIOBase()
                    file.writable = lambda: True
                    file.write = write
                    try:
                        # TextIOWrapper uses this methods to determine
                        # if BOM (for UTF-16, etc) should be added
                        file.seekable = file_or_filename.seekable
                        file.tell = file_or_filename.tell
                    except AttributeError:
                        pass
                file = io.TextIOWrapper(file,
                                        encoding=encoding,
                                        errors="xmlcharrefreplace",
                                        newline="\n")
                # Keep the original file open when the TextIOWrapper is
                # destroyed
                stack.callback(file.detach)
                yield file.write, encoding

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\functions\elementary\miscellaneous.py ===
from sympy.core import S, sympify, NumberKind
from sympy.utilities.iterables import sift
from sympy.core.add import Add
from sympy.core.containers import Tuple
from sympy.core.operations import LatticeOp, ShortCircuit
from sympy.core.function import (Application, Lambda,
    ArgumentIndexError, DefinedFunction)
from sympy.core.expr import Expr
from sympy.core.exprtools import factor_terms
from sympy.core.mod import Mod
from sympy.core.mul import Mul
from sympy.core.numbers import Rational
from sympy.core.power import Pow
from sympy.core.relational import Eq, Relational
from sympy.core.singleton import Singleton
from sympy.core.sorting import ordered
from sympy.core.symbol import Dummy
from sympy.core.rules import Transform
from sympy.core.logic import fuzzy_and, fuzzy_or, _torf
from sympy.core.traversal import walk
from sympy.core.numbers import Integer
from sympy.logic.boolalg import And, Or


def _minmax_as_Piecewise(op, *args):
    # helper for Min/Max rewrite as Piecewise
    from sympy.functions.elementary.piecewise import Piecewise
    ec = []
    for i, a in enumerate(args):
        c = [Relational(a, args[j], op) for j in range(i + 1, len(args))]
        ec.append((a, And(*c)))
    return Piecewise(*ec)


class IdentityFunction(Lambda, metaclass=Singleton):
    """
    The identity function

    Examples
    ========

    >>> from sympy import Id, Symbol
    >>> x = Symbol('x')
    >>> Id(x)
    x

    """

    _symbol = Dummy('x')

    @property
    def signature(self):
        return Tuple(self._symbol)

    @property
    def expr(self):
        return self._symbol


Id = S.IdentityFunction

###############################################################################
############################# ROOT and SQUARE ROOT FUNCTION ###################
###############################################################################


def sqrt(arg, evaluate=None):
    """Returns the principal square root.

    Parameters
    ==========

    evaluate : bool, optional
        The parameter determines if the expression should be evaluated.
        If ``None``, its value is taken from
        ``global_parameters.evaluate``.

    Examples
    ========

    >>> from sympy import sqrt, Symbol, S
    >>> x = Symbol('x')

    >>> sqrt(x)
    sqrt(x)

    >>> sqrt(x)**2
    x

    Note that sqrt(x**2) does not simplify to x.

    >>> sqrt(x**2)
    sqrt(x**2)

    This is because the two are not equal to each other in general.
    For example, consider x == -1:

    >>> from sympy import Eq
    >>> Eq(sqrt(x**2), x).subs(x, -1)
    False

    This is because sqrt computes the principal square root, so the square may
    put the argument in a different branch.  This identity does hold if x is
    positive:

    >>> y = Symbol('y', positive=True)
    >>> sqrt(y**2)
    y

    You can force this simplification by using the powdenest() function with
    the force option set to True:

    >>> from sympy import powdenest
    >>> sqrt(x**2)
    sqrt(x**2)
    >>> powdenest(sqrt(x**2), force=True)
    x

    To get both branches of the square root you can use the rootof function:

    >>> from sympy import rootof

    >>> [rootof(x**2-3,i) for i in (0,1)]
    [-sqrt(3), sqrt(3)]

    Although ``sqrt`` is printed, there is no ``sqrt`` function so looking for
    ``sqrt`` in an expression will fail:

    >>> from sympy.utilities.misc import func_name
    >>> func_name(sqrt(x))
    'Pow'
    >>> sqrt(x).has(sqrt)
    False

    To find ``sqrt`` look for ``Pow`` with an exponent of ``1/2``:

    >>> (x + 1/sqrt(x)).find(lambda i: i.is_Pow and abs(i.exp) is S.Half)
    {1/sqrt(x)}

    See Also
    ========

    sympy.polys.rootoftools.rootof, root, real_root

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Square_root
    .. [2] https://en.wikipedia.org/wiki/Principal_value
    """
    # arg = sympify(arg) is handled by Pow
    return Pow(arg, S.Half, evaluate=evaluate)


def cbrt(arg, evaluate=None):
    """Returns the principal cube root.

    Parameters
    ==========

    evaluate : bool, optional
        The parameter determines if the expression should be evaluated.
        If ``None``, its value is taken from
        ``global_parameters.evaluate``.

    Examples
    ========

    >>> from sympy import cbrt, Symbol
    >>> x = Symbol('x')

    >>> cbrt(x)
    x**(1/3)

    >>> cbrt(x)**3
    x

    Note that cbrt(x**3) does not simplify to x.

    >>> cbrt(x**3)
    (x**3)**(1/3)

    This is because the two are not equal to each other in general.
    For example, consider `x == -1`:

    >>> from sympy import Eq
    >>> Eq(cbrt(x**3), x).subs(x, -1)
    False

    This is because cbrt computes the principal cube root, this
    identity does hold if `x` is positive:

    >>> y = Symbol('y', positive=True)
    >>> cbrt(y**3)
    y

    See Also
    ========

    sympy.polys.rootoftools.rootof, root, real_root

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Cube_root
    .. [2] https://en.wikipedia.org/wiki/Principal_value

    """
    return Pow(arg, Rational(1, 3), evaluate=evaluate)


def root(arg, n, k=0, evaluate=None):
    r"""Returns the *k*-th *n*-th root of ``arg``.

    Parameters
    ==========

    k : int, optional
        Should be an integer in $\{0, 1, ..., n-1\}$.
        Defaults to the principal root if $0$.

    evaluate : bool, optional
        The parameter determines if the expression should be evaluated.
        If ``None``, its value is taken from
        ``global_parameters.evaluate``.

    Examples
    ========

    >>> from sympy import root, Rational
    >>> from sympy.abc import x, n

    >>> root(x, 2)
    sqrt(x)

    >>> root(x, 3)
    x**(1/3)

    >>> root(x, n)
    x**(1/n)

    >>> root(x, -Rational(2, 3))
    x**(-3/2)

    To get the k-th n-th root, specify k:

    >>> root(-2, 3, 2)
    -(-1)**(2/3)*2**(1/3)

    To get all n n-th roots you can use the rootof function.
    The following examples show the roots of unity for n
    equal 2, 3 and 4:

    >>> from sympy import rootof

    >>> [rootof(x**2 - 1, i) for i in range(2)]
    [-1, 1]

    >>> [rootof(x**3 - 1,i) for i in range(3)]
    [1, -1/2 - sqrt(3)*I/2, -1/2 + sqrt(3)*I/2]

    >>> [rootof(x**4 - 1,i) for i in range(4)]
    [-1, 1, -I, I]

    SymPy, like other symbolic algebra systems, returns the
    complex root of negative numbers. This is the principal
    root and differs from the text-book result that one might
    be expecting. For example, the cube root of -8 does not
    come back as -2:

    >>> root(-8, 3)
    2*(-1)**(1/3)

    The real_root function can be used to either make the principal
    result real (or simply to return the real root directly):

    >>> from sympy import real_root
    >>> real_root(_)
    -2
    >>> real_root(-32, 5)
    -2

    Alternatively, the n//2-th n-th root of a negative number can be
    computed with root:

    >>> root(-32, 5, 5//2)
    -2

    See Also
    ========

    sympy.polys.rootoftools.rootof
    sympy.core.intfunc.integer_nthroot
    sqrt, real_root

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Square_root
    .. [2] https://en.wikipedia.org/wiki/Real_root
    .. [3] https://en.wikipedia.org/wiki/Root_of_unity
    .. [4] https://en.wikipedia.org/wiki/Principal_value
    .. [5] https://mathworld.wolfram.com/CubeRoot.html

    """
    n = sympify(n)
    if k:
        return Mul(Pow(arg, S.One/n, evaluate=evaluate), S.NegativeOne**(2*k/n), evaluate=evaluate)
    return Pow(arg, 1/n, evaluate=evaluate)


def real_root(arg, n=None, evaluate=None):
    r"""Return the real *n*'th-root of *arg* if possible.

    Parameters
    ==========

    n : int or None, optional
        If *n* is ``None``, then all instances of
        $(-n)^{1/\text{odd}}$ will be changed to $-n^{1/\text{odd}}$.
        This will only create a real root of a principal root.
        The presence of other factors may cause the result to not be
        real.

    evaluate : bool, optional
        The parameter determines if the expression should be evaluated.
        If ``None``, its value is taken from
        ``global_parameters.evaluate``.

    Examples
    ========

    >>> from sympy import root, real_root

    >>> real_root(-8, 3)
    -2
    >>> root(-8, 3)
    2*(-1)**(1/3)
    >>> real_root(_)
    -2

    If one creates a non-principal root and applies real_root, the
    result will not be real (so use with caution):

    >>> root(-8, 3, 2)
    -2*(-1)**(2/3)
    >>> real_root(_)
    -2*(-1)**(2/3)

    See Also
    ========

    sympy.polys.rootoftools.rootof
    sympy.core.intfunc.integer_nthroot
    root, sqrt
    """
    from sympy.functions.elementary.complexes import Abs, im, sign
    from sympy.functions.elementary.piecewise import Piecewise
    if n is not None:
        return Piecewise(
            (root(arg, n, evaluate=evaluate), Or(Eq(n, S.One), Eq(n, S.NegativeOne))),
            (Mul(sign(arg), root(Abs(arg), n, evaluate=evaluate), evaluate=evaluate),
            And(Eq(im(arg), S.Zero), Eq(Mod(n, 2), S.One))),
            (root(arg, n, evaluate=evaluate), True))
    rv = sympify(arg)
    n1pow = Transform(lambda x: -(-x.base)**x.exp,
                      lambda x:
                      x.is_Pow and
                      x.base.is_negative and
                      x.exp.is_Rational and
                      x.exp.p == 1 and x.exp.q % 2)
    return rv.xreplace(n1pow)

###############################################################################
############################# MINIMUM and MAXIMUM #############################
###############################################################################


class MinMaxBase(Expr, LatticeOp):
    def __new__(cls, *args, **assumptions):
        from sympy.core.parameters import global_parameters
        evaluate = assumptions.pop('evaluate', global_parameters.evaluate)
        args = (sympify(arg) for arg in args)

        # first standard filter, for cls.zero and cls.identity
        # also reshape Max(a, Max(b, c)) to Max(a, b, c)

        if evaluate:
            try:
                args = frozenset(cls._new_args_filter(args))
            except ShortCircuit:
                return cls.zero
            # remove redundant args that are easily identified
            args = cls._collapse_arguments(args, **assumptions)
            # find local zeros
            args = cls._find_localzeros(args, **assumptions)
        args = frozenset(args)

        if not args:
            return cls.identity

        if len(args) == 1:
            return list(args).pop()

        # base creation
        obj = Expr.__new__(cls, *ordered(args), **assumptions)
        obj._argset = args
        return obj

    @classmethod
    def _collapse_arguments(cls, args, **assumptions):
        """Remove redundant args.

        Examples
        ========

        >>> from sympy import Min, Max
        >>> from sympy.abc import a, b, c, d, e

        Any arg in parent that appears in any
        parent-like function in any of the flat args
        of parent can be removed from that sub-arg:

        >>> Min(a, Max(b, Min(a, c, d)))
        Min(a, Max(b, Min(c, d)))

        If the arg of parent appears in an opposite-than parent
        function in any of the flat args of parent that function
        can be replaced with the arg:

        >>> Min(a, Max(b, Min(c, d, Max(a, e))))
        Min(a, Max(b, Min(a, c, d)))
        """
        if not args:
            return args
        args = list(ordered(args))
        if cls == Min:
            other = Max
        else:
            other = Min

        # find global comparable max of Max and min of Min if a new
        # value is being introduced in these args at position 0 of
        # the ordered args
        if args[0].is_number:
            sifted = mins, maxs = [], []
            for i in args:
                for v in walk(i, Min, Max):
                    if v.args[0].is_comparable:
                        sifted[isinstance(v, Max)].append(v)
            small = Min.identity
            for i in mins:
                v = i.args[0]
                if v.is_number and (v < small) == True:
                    small = v
            big = Max.identity
            for i in maxs:
                v = i.args[0]
                if v.is_number and (v > big) == True:
                    big = v
            # at the point when this function is called from __new__,
            # there may be more than one numeric arg present since
            # local zeros have not been handled yet, so look through
            # more than the first arg
            if cls == Min:
                for arg in args:
                    if not arg.is_number:
                        break
                    if (arg < small) == True:
                        small = arg
            elif cls == Max:
                for arg in args:
                    if not arg.is_number:
                        break
                    if (arg > big) == True:
                        big = arg
            T = None
            if cls == Min:
                if small != Min.identity:
                    other = Max
                    T = small
            elif big != Max.identity:
                other = Min
                T = big
            if T is not None:
                # remove numerical redundancy
                for i in range(len(args)):
                    a = args[i]
                    if isinstance(a, other):
                        a0 = a.args[0]
                        if ((a0 > T) if other == Max else (a0 < T)) == True:
                            args[i] = cls.identity

        # remove redundant symbolic args
        def do(ai, a):
            if not isinstance(ai, (Min, Max)):
                return ai
            cond = a in ai.args
            if not cond:
                return ai.func(*[do(i, a) for i in ai.args],
                    evaluate=False)
            if isinstance(ai, cls):
                return ai.func(*[do(i, a) for i in ai.args if i != a],
                    evaluate=False)
            return a
        for i, a in enumerate(args):
            args[i + 1:] = [do(ai, a) for ai in args[i + 1:]]

        # factor out common elements as for
        # Min(Max(x, y), Max(x, z)) -> Max(x, Min(y, z))
        # and vice versa when swapping Min/Max -- do this only for the
        # easy case where all functions contain something in common;
        # trying to find some optimal subset of args to modify takes
        # too long

        def factor_minmax(args):
            is_other = lambda arg: isinstance(arg, other)
            other_args, remaining_args = sift(args, is_other, binary=True)
            if not other_args:
                return args

            # Min(Max(x, y, z), Max(x, y, u, v)) -> {x,y}, ({z}, {u,v})
            arg_sets = [set(arg.args) for arg in other_args]
            common = set.intersection(*arg_sets)
            if not common:
                return args

            new_other_args = list(common)
            arg_sets_diff = [arg_set - common for arg_set in arg_sets]

            # If any set is empty after removing common then all can be
            # discarded e.g. Min(Max(a, b, c), Max(a, b)) -> Max(a, b)
            if all(arg_sets_diff):
                other_args_diff = [other(*s, evaluate=False) for s in arg_sets_diff]
                new_other_args.append(cls(*other_args_diff, evaluate=False))

            other_args_factored = other(*new_other_args, evaluate=False)
            return remaining_args + [other_args_factored]

        if len(args) > 1:
            args = factor_minmax(args)

        return args

    @classmethod
    def _new_args_filter(cls, arg_sequence):
        """
        Generator filtering args.

        first standard filter, for cls.zero and cls.identity.
        Also reshape ``Max(a, Max(b, c))`` to ``Max(a, b, c)``,
        and check arguments for comparability
        """
        for arg in arg_sequence:
            # pre-filter, checking comparability of arguments
            if not isinstance(arg, Expr) or arg.is_extended_real is False or (
                    arg.is_number and
                    not arg.is_comparable):
                raise ValueError("The argument '%s' is not comparable." % arg)

            if arg == cls.zero:
                raise ShortCircuit(arg)
            elif arg == cls.identity:
                continue
            elif arg.func == cls:
                yield from arg.args
            else:
                yield arg

    @classmethod
    def _find_localzeros(cls, values, **options):
        """
        Sequentially allocate values to localzeros.

        When a value is identified as being more extreme than another member it
        replaces that member; if this is never true, then the value is simply
        appended to the localzeros.
        """
        localzeros = set()
        for v in values:
            is_newzero = True
            localzeros_ = list(localzeros)
            for z in localzeros_:
                if id(v) == id(z):
                    is_newzero = False
                else:
                    con = cls._is_connected(v, z)
                    if con:
                        is_newzero = False
                        if con is True or con == cls:
                            localzeros.remove(z)
                            localzeros.update([v])
            if is_newzero:
                localzeros.update([v])
        return localzeros

    @classmethod
    def _is_connected(cls, x, y):
        """
        Check if x and y are connected somehow.
        """
        for i in range(2):
            if x == y:
                return True
            t, f = Max, Min
            for op in "><":
                for j in range(2):
                    try:
                        if op == ">":
                            v = x >= y
                        else:
                            v = x <= y
                    except TypeError:
                        return False  # non-real arg
                    if not v.is_Relational:
                        return t if v else f
                    t, f = f, t
                    x, y = y, x
                x, y = y, x  # run next pass with reversed order relative to start
            # simplification can be expensive, so be conservative
            # in what is attempted
            x = factor_terms(x - y)
            y = S.Zero

        return False

    def _eval_derivative(self, s):
        # f(x).diff(s) -> x.diff(s) * f.fdiff(1)(s)
        i = 0
        l = []
        for a in self.args:
            i += 1
            da = a.diff(s)
            if da.is_zero:
                continue
            try:
                df = self.fdiff(i)
            except ArgumentIndexError:
                df = super().fdiff(i)
            l.append(df * da)
        return Add(*l)

    def _eval_rewrite_as_Abs(self, *args, **kwargs):
        from sympy.functions.elementary.complexes import Abs
        s = (args[0] + self.func(*args[1:]))/2
        d = abs(args[0] - self.func(*args[1:]))/2
        return (s + d if isinstance(self, Max) else s - d).rewrite(Abs)

    def evalf(self, n=15, **options):
        return self.func(*[a.evalf(n, **options) for a in self.args])

    def n(self, *args, **kwargs):
        return self.evalf(*args, **kwargs)

    _eval_is_algebraic = lambda s: _torf(i.is_algebraic for i in s.args)
    _eval_is_antihermitian = lambda s: _torf(i.is_antihermitian for i in s.args)
    _eval_is_commutative = lambda s: _torf(i.is_commutative for i in s.args)
    _eval_is_complex = lambda s: _torf(i.is_complex for i in s.args)
    _eval_is_composite = lambda s: _torf(i.is_composite for i in s.args)
    _eval_is_even = lambda s: _torf(i.is_even for i in s.args)
    _eval_is_finite = lambda s: _torf(i.is_finite for i in s.args)
    _eval_is_hermitian = lambda s: _torf(i.is_hermitian for i in s.args)
    _eval_is_imaginary = lambda s: _torf(i.is_imaginary for i in s.args)
    _eval_is_infinite = lambda s: _torf(i.is_infinite for i in s.args)
    _eval_is_integer = lambda s: _torf(i.is_integer for i in s.args)
    _eval_is_irrational = lambda s: _torf(i.is_irrational for i in s.args)
    _eval_is_negative = lambda s: _torf(i.is_negative for i in s.args)
    _eval_is_noninteger = lambda s: _torf(i.is_noninteger for i in s.args)
    _eval_is_nonnegative = lambda s: _torf(i.is_nonnegative for i in s.args)
    _eval_is_nonpositive = lambda s: _torf(i.is_nonpositive for i in s.args)
    _eval_is_nonzero = lambda s: _torf(i.is_nonzero for i in s.args)
    _eval_is_odd = lambda s: _torf(i.is_odd for i in s.args)
    _eval_is_polar = lambda s: _torf(i.is_polar for i in s.args)
    _eval_is_positive = lambda s: _torf(i.is_positive for i in s.args)
    _eval_is_prime = lambda s: _torf(i.is_prime for i in s.args)
    _eval_is_rational = lambda s: _torf(i.is_rational for i in s.args)
    _eval_is_real = lambda s: _torf(i.is_real for i in s.args)
    _eval_is_extended_real = lambda s: _torf(i.is_extended_real for i in s.args)
    _eval_is_transcendental = lambda s: _torf(i.is_transcendental for i in s.args)
    _eval_is_zero = lambda s: _torf(i.is_zero for i in s.args)


class Max(MinMaxBase, Application):
    r"""
    Return, if possible, the maximum value of the list.

    When number of arguments is equal one, then
    return this argument.

    When number of arguments is equal two, then
    return, if possible, the value from (a, b) that is $\ge$ the other.

    In common case, when the length of list greater than 2, the task
    is more complicated. Return only the arguments, which are greater
    than others, if it is possible to determine directional relation.

    If is not possible to determine such a relation, return a partially
    evaluated result.

    Assumptions are used to make the decision too.

    Also, only comparable arguments are permitted.

    It is named ``Max`` and not ``max`` to avoid conflicts
    with the built-in function ``max``.


    Examples
    ========

    >>> from sympy import Max, Symbol, oo
    >>> from sympy.abc import x, y, z
    >>> p = Symbol('p', positive=True)
    >>> n = Symbol('n', negative=True)

    >>> Max(x, -2)
    Max(-2, x)
    >>> Max(x, -2).subs(x, 3)
    3
    >>> Max(p, -2)
    p
    >>> Max(x, y)
    Max(x, y)
    >>> Max(x, y) == Max(y, x)
    True
    >>> Max(x, Max(y, z))
    Max(x, y, z)
    >>> Max(n, 8, p, 7, -oo)
    Max(8, p)
    >>> Max (1, x, oo)
    oo

    * Algorithm

    The task can be considered as searching of supremums in the
    directed complete partial orders [1]_.

    The source values are sequentially allocated by the isolated subsets
    in which supremums are searched and result as Max arguments.

    If the resulted supremum is single, then it is returned.

    The isolated subsets are the sets of values which are only the comparable
    with each other in the current set. E.g. natural numbers are comparable with
    each other, but not comparable with the `x` symbol. Another example: the
    symbol `x` with negative assumption is comparable with a natural number.

    Also there are "least" elements, which are comparable with all others,
    and have a zero property (maximum or minimum for all elements).
    For example, in case of $\infty$, the allocation operation is terminated
    and only this value is returned.

    Assumption:
       - if $A > B > C$ then $A > C$
       - if $A = B$ then $B$ can be removed

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Directed_complete_partial_order
    .. [2] https://en.wikipedia.org/wiki/Lattice_%28order%29

    See Also
    ========

    Min : find minimum values
    """
    zero = S.Infinity
    identity = S.NegativeInfinity

    def fdiff( self, argindex ):
        from sympy.functions.special.delta_functions import Heaviside
        n = len(self.args)
        if 0 < argindex and argindex <= n:
            argindex -= 1
            if n == 2:
                return Heaviside(self.args[argindex] - self.args[1 - argindex])
            newargs = tuple([self.args[i] for i in range(n) if i != argindex])
            return Heaviside(self.args[argindex] - Max(*newargs))
        else:
            raise ArgumentIndexError(self, argindex)

    def _eval_rewrite_as_Heaviside(self, *args, **kwargs):
        from sympy.functions.special.delta_functions import Heaviside
        return Add(*[j*Mul(*[Heaviside(j - i) for i in args if i!=j]) \
                for j in args])

    def _eval_rewrite_as_Piecewise(self, *args, **kwargs):
        return _minmax_as_Piecewise('>=', *args)

    def _eval_is_positive(self):
        return fuzzy_or(a.is_positive for a in self.args)

    def _eval_is_nonnegative(self):
        return fuzzy_or(a.is_nonnegative for a in self.args)

    def _eval_is_negative(self):
        return fuzzy_and(a.is_negative for a in self.args)


class Min(MinMaxBase, Application):
    """
    Return, if possible, the minimum value of the list.
    It is named ``Min`` and not ``min`` to avoid conflicts
    with the built-in function ``min``.

    Examples
    ========

    >>> from sympy import Min, Symbol, oo
    >>> from sympy.abc import x, y
    >>> p = Symbol('p', positive=True)
    >>> n = Symbol('n', negative=True)

    >>> Min(x, -2)
    Min(-2, x)
    >>> Min(x, -2).subs(x, 3)
    -2
    >>> Min(p, -3)
    -3
    >>> Min(x, y)
    Min(x, y)
    >>> Min(n, 8, p, -7, p, oo)
    Min(-7, n)

    See Also
    ========

    Max : find maximum values
    """
    zero = S.NegativeInfinity
    identity = S.Infinity

    def fdiff( self, argindex ):
        from sympy.functions.special.delta_functions import Heaviside
        n = len(self.args)
        if 0 < argindex and argindex <= n:
            argindex -= 1
            if n == 2:
                return Heaviside( self.args[1-argindex] - self.args[argindex] )
            newargs = tuple([ self.args[i] for i in range(n) if i != argindex])
            return Heaviside( Min(*newargs) - self.args[argindex] )
        else:
            raise ArgumentIndexError(self, argindex)

    def _eval_rewrite_as_Heaviside(self, *args, **kwargs):
        from sympy.functions.special.delta_functions import Heaviside
        return Add(*[j*Mul(*[Heaviside(i-j) for i in args if i!=j]) \
                for j in args])

    def _eval_rewrite_as_Piecewise(self, *args, **kwargs):
        return _minmax_as_Piecewise('<=', *args)

    def _eval_is_positive(self):
        return fuzzy_and(a.is_positive for a in self.args)

    def _eval_is_nonnegative(self):
        return fuzzy_and(a.is_nonnegative for a in self.args)

    def _eval_is_negative(self):
        return fuzzy_or(a.is_negative for a in self.args)


class Rem(DefinedFunction):
    """Returns the remainder when ``p`` is divided by ``q`` where ``p`` is finite
    and ``q`` is not equal to zero. The result, ``p - int(p/q)*q``, has the same sign
    as the divisor.

    Parameters
    ==========

    p : Expr
        Dividend.

    q : Expr
        Divisor.

    Notes
    =====

    ``Rem`` corresponds to the ``%`` operator in C.

    Examples
    ========

    >>> from sympy.abc import x, y
    >>> from sympy import Rem
    >>> Rem(x**3, y)
    Rem(x**3, y)
    >>> Rem(x**3, y).subs({x: -5, y: 3})
    -2

    See Also
    ========

    Mod
    """
    kind = NumberKind

    @classmethod
    def eval(cls, p, q):
        """Return the function remainder if both p, q are numbers and q is not
        zero.
        """

        if q.is_zero:
            raise ZeroDivisionError("Division by zero")
        if p is S.NaN or q is S.NaN or p.is_finite is False or q.is_finite is False:
            return S.NaN
        if p is S.Zero or p in (q, -q) or (p.is_integer and q == 1):
            return S.Zero

        if q.is_Number:
            if p.is_Number:
                return p - Integer(p/q)*q

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\logic\algorithms\lra_theory.py ===
"""Implements "A Fast Linear-Arithmetic Solver for DPLL(T)"

The LRASolver class defined in this file can be used
in conjunction with a SAT solver to check the
satisfiability of formulas involving inequalities.

Here's an example of how that would work:

    Suppose you want to check the satisfiability of
    the following formula:

    >>> from sympy.core.relational import Eq
    >>> from sympy.abc import x, y
    >>> f = ((x > 0) | (x < 0)) & (Eq(x, 0) | Eq(y, 1)) & (~Eq(y, 1) | Eq(1, 2))

    First a preprocessing step should be done on f. During preprocessing,
    f should be checked for any predicates such as `Q.prime` that can't be
    handled. Also unequality like `~Eq(y, 1)` should be split.

    I should mention that the paper says to split both equalities and
    unequality, but this implementation only requires that unequality
    be split.

    >>> f = ((x > 0) | (x < 0)) & (Eq(x, 0) | Eq(y, 1)) & ((y < 1) | (y > 1) | Eq(1, 2))

    Then an LRASolver instance needs to be initialized with this formula.

    >>> from sympy.assumptions.cnf import CNF, EncodedCNF
    >>> from sympy.assumptions.ask import Q
    >>> from sympy.logic.algorithms.lra_theory import LRASolver
    >>> cnf = CNF.from_prop(f)
    >>> enc = EncodedCNF()
    >>> enc.add_from_cnf(cnf)
    >>> lra, conflicts = LRASolver.from_encoded_cnf(enc)

    Any immediate one-lital conflicts clauses will be detected here.
    In this example, `~Eq(1, 2)` is one such conflict clause. We'll
    want to add it to `f` so that the SAT solver is forced to
    assign Eq(1, 2) to False.

    >>> f = f & ~Eq(1, 2)

    Now that the one-literal conflict clauses have been added
    and an lra object has been initialized, we can pass `f`
    to a SAT solver. The SAT solver will give us a satisfying
    assignment such as:

    (1 = 2): False
    (y = 1): True
    (y < 1): True
    (y > 1): True
    (x = 0): True
    (x < 0): True
    (x > 0): True

    Next you would pass this assignment to the LRASolver
    which will be able to determine that this particular
    assignment is satisfiable or not.

    Note that since EncodedCNF is inherently non-deterministic,
    the int each predicate is encoded as is not consistent. As a
    result, the code below likely does not reflect the assignment
    given above.

    >>> lra.assert_lit(-1) #doctest: +SKIP
    >>> lra.assert_lit(2) #doctest: +SKIP
    >>> lra.assert_lit(3) #doctest: +SKIP
    >>> lra.assert_lit(4) #doctest: +SKIP
    >>> lra.assert_lit(5) #doctest: +SKIP
    >>> lra.assert_lit(6) #doctest: +SKIP
    >>> lra.assert_lit(7) #doctest: +SKIP
    >>> is_sat, conflict_or_assignment = lra.check()

    As the particular assignment suggested is not satisfiable,
    the LRASolver will return unsat and a conflict clause when
    given that assignment. The conflict clause will always be
    minimal, but there can be multiple minimal conflict clauses.
    One possible conflict clause could be `~(x < 0) | ~(x > 0)`.

    We would then add whatever conflict clause is given to
    `f` to prevent the SAT solver from coming up with an
    assignment with the same conflicting literals. In this case,
    the conflict clause `~(x < 0) | ~(x > 0)` would prevent
    any assignment where both (x < 0) and (x > 0) were both
    true.

    The SAT solver would then find another assignment
    and we would check that assignment with the LRASolver
    and so on. Eventually either a satisfying assignment
    that the SAT solver and LRASolver agreed on would be found
    or enough conflict clauses would be added so that the
    boolean formula was unsatisfiable.


This implementation is based on [1]_, which includes a
detailed explanation of the algorithm and pseudocode
for the most important functions.

[1]_ also explains how backtracking and theory propagation
could be implemented to speed up the current implementation,
but these are not currently implemented.

TODO:
 - Handle non-rational real numbers
 - Handle positive and negative infinity
 - Implement backtracking and theory proposition
 - Simplify matrix by removing unused variables using Gaussian elimination

References
==========

.. [1] Dutertre, B., de Moura, L.:
       A Fast Linear-Arithmetic Solver for DPLL(T)
       https://link.springer.com/chapter/10.1007/11817963_11
"""
from sympy.solvers.solveset import linear_eq_to_matrix
from sympy.matrices.dense import eye
from sympy.assumptions import Predicate
from sympy.assumptions.assume import AppliedPredicate
from sympy.assumptions.ask import Q
from sympy.core import Dummy
from sympy.core.mul import Mul
from sympy.core.add import Add
from sympy.core.relational import Eq, Ne
from sympy.core.sympify import sympify
from sympy.core.singleton import S
from sympy.core.numbers import Rational, oo
from sympy.matrices.dense import Matrix

class UnhandledInput(Exception):
    """
    Raised while creating an LRASolver if non-linearity
    or non-rational numbers are present.
    """

# predicates that LRASolver understands and makes use of
ALLOWED_PRED = {Q.eq, Q.gt, Q.lt, Q.le, Q.ge}

# if true ~Q.gt(x, y) implies Q.le(x, y)
HANDLE_NEGATION = True

class LRASolver():
    """
    Linear Arithmetic Solver for DPLL(T) implemented with an algorithm based on
    the Dual Simplex method. Uses Bland's pivoting rule to avoid cycling.

    References
    ==========

    .. [1] Dutertre, B., de Moura, L.:
           A Fast Linear-Arithmetic Solver for DPLL(T)
           https://link.springer.com/chapter/10.1007/11817963_11
    """

    def __init__(self, A, slack_variables, nonslack_variables, enc_to_boundary, s_subs, testing_mode):
        """
        Use the "from_encoded_cnf" method to create a new LRASolver.
        """
        self.run_checks = testing_mode
        self.s_subs = s_subs  # used only for test_lra_theory.test_random_problems

        if any(not isinstance(a, Rational) for a in A):
            raise UnhandledInput("Non-rational numbers are not handled")
        if any(not isinstance(b.bound, Rational) for b in enc_to_boundary.values()):
            raise UnhandledInput("Non-rational numbers are not handled")
        m, n = len(slack_variables), len(slack_variables)+len(nonslack_variables)
        if m != 0:
            assert A.shape == (m, n)
        if self.run_checks:
            assert A[:, n-m:] == -eye(m)

        self.enc_to_boundary = enc_to_boundary  # mapping of int to Boundary objects
        self.boundary_to_enc = {value: key for key, value in enc_to_boundary.items()}
        self.A = A
        self.slack = slack_variables
        self.nonslack = nonslack_variables
        self.all_var = nonslack_variables + slack_variables

        self.slack_set = set(slack_variables)

        self.is_sat = True  # While True, all constraints asserted so far are satisfiable
        self.result = None  # always one of: (True, assignment), (False, conflict clause), None

    @staticmethod
    def from_encoded_cnf(encoded_cnf, testing_mode=False):
        """
        Creates an LRASolver from an EncodedCNF object
        and a list of conflict clauses for propositions
        that can be simplified to True or False.

        Parameters
        ==========

        encoded_cnf : EncodedCNF

        testing_mode : bool
            Setting testing_mode to True enables some slow assert statements
            and sorting to reduce nonterministic behavior.

        Returns
        =======

        (lra, conflicts)

        lra : LRASolver

        conflicts : list
            Contains a one-literal conflict clause for each proposition
            that can be simplified to True or False.

        Example
        =======

        >>> from sympy.core.relational import Eq
        >>> from sympy.assumptions.cnf import CNF, EncodedCNF
        >>> from sympy.assumptions.ask import Q
        >>> from sympy.logic.algorithms.lra_theory import LRASolver
        >>> from sympy.abc import x, y, z
        >>> phi = (x >= 0) & ((x + y <= 2) | (x + 2 * y - z >= 6))
        >>> phi = phi & (Eq(x + y, 2) | (x + 2 * y - z > 4))
        >>> phi = phi & Q.gt(2, 1)
        >>> cnf = CNF.from_prop(phi)
        >>> enc = EncodedCNF()
        >>> enc.from_cnf(cnf)
        >>> lra, conflicts = LRASolver.from_encoded_cnf(enc, testing_mode=True)
        >>> lra #doctest: +SKIP
        <sympy.logic.algorithms.lra_theory.LRASolver object at 0x7fdcb0e15b70>
        >>> conflicts #doctest: +SKIP
        [[4]]
        """
        # This function has three main jobs:
        # - raise errors if the input formula is not handled
        # - preprocesses the formula into a matrix and single variable constraints
        # - create one-literal conflict clauses from predicates that are always True
        #   or always False such as Q.gt(3, 2)
        #
        # See the preprocessing section of "A Fast Linear-Arithmetic Solver for DPLL(T)"
        # for an explanation of how the formula is converted into a matrix
        # and a set of single variable constraints.

        encoding = {}  # maps int to boundary
        A = []

        basic = []
        s_count = 0
        s_subs = {}
        nonbasic = []

        if testing_mode:
            # sort to reduce nondeterminism
            encoded_cnf_items = sorted(encoded_cnf.encoding.items(), key=lambda x: str(x))
        else:
            encoded_cnf_items = encoded_cnf.encoding.items()

        empty_var = Dummy()
        var_to_lra_var = {}
        conflicts = []

        for prop, enc in encoded_cnf_items:
            if isinstance(prop, Predicate):
                prop = prop(empty_var)
            if not isinstance(prop, AppliedPredicate):
                if prop == True:
                    conflicts.append([enc])
                    continue
                if prop == False:
                    conflicts.append([-enc])
                    continue

                raise ValueError(f"Unhandled Predicate: {prop}")

            assert prop.function in ALLOWED_PRED
            if prop.lhs == S.NaN or prop.rhs == S.NaN:
                raise ValueError(f"{prop} contains nan")
            if prop.lhs.is_imaginary or prop.rhs.is_imaginary:
                raise UnhandledInput(f"{prop} contains an imaginary component")
            if prop.lhs == oo or prop.rhs == oo:
                raise UnhandledInput(f"{prop} contains infinity")

            prop = _eval_binrel(prop)  # simplify variable-less quantities to True / False if possible
            if prop == True:
                conflicts.append([enc])
                continue
            elif prop == False:
                conflicts.append([-enc])
                continue
            elif prop is None:
                raise UnhandledInput(f"{prop} could not be simplified")

            expr = prop.lhs - prop.rhs
            if prop.function in [Q.ge, Q.gt]:
                expr = -expr

            # expr should be less than (or equal to) 0
            # otherwise prop is False
            if prop.function in [Q.le, Q.ge]:
                bool = (expr <= 0)
            elif prop.function in [Q.lt, Q.gt]:
                bool = (expr < 0)
            else:
                assert prop.function == Q.eq
                bool = Eq(expr, 0)

            if bool == True:
                conflicts.append([enc])
                continue
            elif bool == False:
                conflicts.append([-enc])
                continue


            vars, const = _sep_const_terms(expr)  # example: (2x + 3y + 2) --> (2x + 3y), (2)
            vars, var_coeff = _sep_const_coeff(vars)  # examples: (2x) --> (x, 2); (2x + 3y) --> (2x + 3y), (1)
            const = const / var_coeff

            terms = _list_terms(vars)  # example: (2x + 3y) --> [2x, 3y]
            for term in terms:
                term, _ = _sep_const_coeff(term)
                assert len(term.free_symbols) > 0
                if term not in var_to_lra_var:
                    var_to_lra_var[term] = LRAVariable(term)
                    nonbasic.append(term)

            if len(terms) > 1:
                if vars not in s_subs:
                    s_count += 1
                    d = Dummy(f"s{s_count}")
                    var_to_lra_var[d] = LRAVariable(d)
                    basic.append(d)
                    s_subs[vars] = d
                    A.append(vars - d)
                var = s_subs[vars]
            else:
                var = terms[0]

            assert var_coeff != 0

            equality = prop.function == Q.eq
            upper = var_coeff > 0 if not equality else None
            strict = prop.function in [Q.gt, Q.lt]
            b = Boundary(var_to_lra_var[var], -const, upper, equality, strict)
            encoding[enc] = b

        fs = [v.free_symbols for v in nonbasic + basic]
        assert all(len(syms) > 0 for syms in fs)
        fs_count = sum(len(syms) for syms in fs)
        if len(fs) > 0 and  len(set.union(*fs)) < fs_count:
            raise UnhandledInput("Nonlinearity is not handled")

        A, _ = linear_eq_to_matrix(A, nonbasic + basic)
        nonbasic = [var_to_lra_var[nb] for nb in nonbasic]
        basic = [var_to_lra_var[b] for b in basic]
        for idx, var in enumerate(nonbasic + basic):
            var.col_idx = idx

        return LRASolver(A, basic, nonbasic, encoding, s_subs, testing_mode), conflicts

    def reset_bounds(self):
        """
        Resets the state of the LRASolver to before
        anything was asserted.
        """
        self.result = None
        for var in self.all_var:
            var.lower = LRARational(-float("inf"), 0)
            var.lower_from_eq = False
            var.lower_from_neg = False
            var.upper = LRARational(float("inf"), 0)
            var.upper_from_eq= False
            var.lower_from_neg = False
            var.assign = LRARational(0, 0)

    def assert_lit(self, enc_constraint):
        """
        Assert a literal representing a constraint
        and update the internal state accordingly.

        Note that due to peculiarities of this implementation
        asserting ~(x > 0) will assert (x <= 0) but asserting
        ~Eq(x, 0) will not do anything.

        Parameters
        ==========

        enc_constraint : int
            A mapping of encodings to constraints
            can be found in `self.enc_to_boundary`.

        Returns
        =======

        None or (False, explanation)

        explanation : set of ints
            A conflict clause that "explains" why
            the literals asserted so far are unsatisfiable.
        """
        if abs(enc_constraint) not in self.enc_to_boundary:
            return None

        if not HANDLE_NEGATION and enc_constraint < 0:
            return None

        boundary = self.enc_to_boundary[abs(enc_constraint)]
        sym, c, negated = boundary.var, boundary.bound, enc_constraint < 0

        if boundary.equality and negated:
            return None # negated equality is not handled and should only appear in conflict clauses

        upper = boundary.upper != negated
        if boundary.strict != negated:
            delta = -1 if upper else 1
            c = LRARational(c, delta)
        else:
            c = LRARational(c, 0)

        if boundary.equality:
            res1 = self._assert_lower(sym, c, from_equality=True, from_neg=negated)
            if res1 and res1[0] == False:
                res = res1
            else:
                res2 = self._assert_upper(sym, c, from_equality=True, from_neg=negated)
                res =  res2
        elif upper:
            res = self._assert_upper(sym, c, from_neg=negated)
        else:
            res = self._assert_lower(sym, c, from_neg=negated)

        if self.is_sat and sym not in self.slack_set:
            self.is_sat = res is None
        else:
            self.is_sat = False

        return res

    def _assert_upper(self, xi, ci, from_equality=False, from_neg=False):
        """
        Adjusts the upper bound on variable xi if the new upper bound is
        more limiting. The assignment of variable xi is adjusted to be
        within the new bound if needed.

        Also calls `self._update` to update the assignment for slack variables
        to keep all equalities satisfied.
        """
        if self.result:
            assert self.result[0] != False
        self.result = None
        if ci >= xi.upper:
            return None
        if ci < xi.lower:
            assert (xi.lower[1] >= 0) is True
            assert (ci[1] <= 0) is True

            lit1, neg1 = Boundary.from_lower(xi)

            lit2 = Boundary(var=xi, const=ci[0], strict=ci[1] != 0, upper=True, equality=from_equality)
            if from_neg:
                lit2 = lit2.get_negated()
            neg2 = -1 if from_neg else 1

            conflict = [-neg1*self.boundary_to_enc[lit1], -neg2*self.boundary_to_enc[lit2]]
            self.result = False, conflict
            return self.result
        xi.upper = ci
        xi.upper_from_eq = from_equality
        xi.upper_from_neg = from_neg
        if xi in self.nonslack and xi.assign > ci:
            self._update(xi, ci)

        if self.run_checks and all(v.assign[0] != float("inf") and v.assign[0] != -float("inf")
                                   for v in self.all_var):
            M = self.A
            X = Matrix([v.assign[0] for v in self.all_var])
            assert all(abs(val) < 10 ** (-10) for val in M * X)

        return None

    def _assert_lower(self, xi, ci, from_equality=False, from_neg=False):
        """
        Adjusts the lower bound on variable xi if the new lower bound is
        more limiting. The assignment of variable xi is adjusted to be
        within the new bound if needed.

        Also calls `self._update` to update the assignment for slack variables
        to keep all equalities satisfied.
        """
        if self.result:
            assert self.result[0] != False
        self.result = None
        if ci <= xi.lower:
            return None
        if ci > xi.upper:
            assert (xi.upper[1] <= 0) is True
            assert (ci[1] >= 0) is True

            lit1, neg1 = Boundary.from_upper(xi)

            lit2 = Boundary(var=xi, const=ci[0], strict=ci[1] != 0, upper=False, equality=from_equality)
            if from_neg:
                lit2 = lit2.get_negated()
            neg2 = -1 if from_neg else 1

            conflict = [-neg1*self.boundary_to_enc[lit1],-neg2*self.boundary_to_enc[lit2]]
            self.result = False, conflict
            return self.result
        xi.lower = ci
        xi.lower_from_eq = from_equality
        xi.lower_from_neg = from_neg
        if xi in self.nonslack and xi.assign < ci:
            self._update(xi, ci)

        if self.run_checks and all(v.assign[0] != float("inf") and v.assign[0] != -float("inf")
                                   for v in self.all_var):
            M = self.A
            X = Matrix([v.assign[0] for v in self.all_var])
            assert all(abs(val) < 10 ** (-10) for val in M * X)

        return None

    def _update(self, xi, v):
        """
        Updates all slack variables that have equations that contain
        variable xi so that they stay satisfied given xi is equal to v.
        """
        i = xi.col_idx
        for j, b in enumerate(self.slack):
            aji = self.A[j, i]
            b.assign = b.assign + (v - xi.assign)*aji
        xi.assign = v

    def check(self):
        """
        Searches for an assignment that satisfies all constraints
        or determines that no such assignment exists and gives
        a minimal conflict clause that "explains" why the
        constraints are unsatisfiable.

        Returns
        =======

        (True, assignment) or (False, explanation)

        assignment : dict of LRAVariables to values
            Assigned values are tuples that represent a rational number
            plus some infinatesimal delta.

        explanation : set of ints
        """
        if self.is_sat:
            return True, {var: var.assign for var in self.all_var}
        if self.result:
            return self.result

        from sympy.matrices.dense import Matrix
        M = self.A.copy()
        basic = {s: i for i, s in enumerate(self.slack)}  # contains the row index associated with each basic variable
        nonbasic = set(self.nonslack)
        while True:
            if self.run_checks:
                # nonbasic variables must always be within bounds
                assert all(((nb.assign >= nb.lower) == True) and ((nb.assign <= nb.upper) == True) for nb in nonbasic)

                # assignments for x must always satisfy Ax = 0
                # probably have to turn this off when dealing with strict ineq
                if all(v.assign[0] != float("inf") and v.assign[0] != -float("inf")
                                   for v in self.all_var):
                    X = Matrix([v.assign[0] for v in self.all_var])
                    assert all(abs(val) < 10**(-10) for val in M*X)

                # check upper and lower match this format:
                # x <= rat + delta iff x < rat
                # x >= rat - delta iff x > rat
                # this wouldn't make sense:
                # x <= rat - delta
                # x >= rat + delta
                assert all(x.upper[1] <= 0 for x in self.all_var)
                assert all(x.lower[1] >= 0 for x in self.all_var)

            cand = [b for b in basic if b.assign < b.lower or b.assign > b.upper]

            if len(cand) == 0:
                return True, {var: var.assign for var in self.all_var}

            xi = min(cand, key=lambda v: v.col_idx) # Bland's rule
            i = basic[xi]

            if xi.assign < xi.lower:
                cand = [nb for nb in nonbasic
                        if (M[i, nb.col_idx] > 0 and nb.assign < nb.upper)
                        or (M[i, nb.col_idx] < 0 and nb.assign > nb.lower)]
                if len(cand) == 0:
                    N_plus = [nb for nb in nonbasic if M[i, nb.col_idx] > 0]
                    N_minus = [nb for nb in nonbasic if M[i, nb.col_idx] < 0]

                    conflict = []
                    conflict += [Boundary.from_upper(nb) for nb in N_plus]
                    conflict += [Boundary.from_lower(nb) for nb in N_minus]
                    conflict.append(Boundary.from_lower(xi))
                    conflict = [-neg*self.boundary_to_enc[c] for c, neg in conflict]
                    return False, conflict
                xj = min(cand, key=str)
                M = self._pivot_and_update(M, basic, nonbasic, xi, xj, xi.lower)

            if xi.assign > xi.upper:
                cand = [nb for nb in nonbasic
                        if (M[i, nb.col_idx] < 0 and nb.assign < nb.upper)
                        or (M[i, nb.col_idx] > 0 and nb.assign > nb.lower)]

                if len(cand) == 0:
                    N_plus = [nb for nb in nonbasic if M[i, nb.col_idx] > 0]
                    N_minus = [nb for nb in nonbasic if M[i, nb.col_idx] < 0]

                    conflict = []
                    conflict += [Boundary.from_upper(nb) for nb in N_minus]
                    conflict += [Boundary.from_lower(nb) for nb in N_plus]
                    conflict.append(Boundary.from_upper(xi))

                    conflict = [-neg*self.boundary_to_enc[c] for c, neg in conflict]
                    return False, conflict
                xj = min(cand, key=lambda v: v.col_idx)
                M = self._pivot_and_update(M, basic, nonbasic, xi, xj, xi.upper)

    def _pivot_and_update(self, M, basic, nonbasic, xi, xj, v):
        """
        Pivots basic variable xi with nonbasic variable xj,
        and sets value of xi to v and adjusts the values of all basic variables
        to keep equations satisfied.
        """
        i, j = basic[xi], xj.col_idx
        assert M[i, j] != 0
        theta = (v - xi.assign)*(1/M[i, j])
        xi.assign = v
        xj.assign = xj.assign + theta
        for xk in basic:
            if xk != xi:
                k = basic[xk]
                akj = M[k, j]
                xk.assign = xk.assign + theta*akj
        # pivot
        basic[xj] = basic[xi]
        del basic[xi]
        nonbasic.add(xi)
        nonbasic.remove(xj)
        return self._pivot(M, i, j)

    @staticmethod
    def _pivot(M, i, j):
        """
        Performs a pivot operation about entry i, j of M by performing
        a series of row operations on a copy of M and returning the result.
        The original M is left unmodified.

        Conceptually, M represents a system of equations and pivoting
        can be thought of as rearranging equation i to be in terms of
        variable j and then substituting in the rest of the equations
        to get rid of other occurances of variable j.

        Example
        =======

        >>> from sympy.matrices.dense import Matrix
        >>> from sympy.logic.algorithms.lra_theory import LRASolver
        >>> from sympy import var
        >>> Matrix(3, 3, var('a:i'))
        Matrix([
        [a, b, c],
        [d, e, f],
        [g, h, i]])

        This matrix is equivalent to:
        0 = a*x + b*y + c*z
        0 = d*x + e*y + f*z
        0 = g*x + h*y + i*z

        >>> LRASolver._pivot(_, 1, 0)
        Matrix([
        [ 0, -a*e/d + b, -a*f/d + c],
        [-1,       -e/d,       -f/d],
        [ 0,  h - e*g/d,  i - f*g/d]])

        We rearrange equation 1 in terms of variable 0 (x)
        and substitute to remove x from the other equations.

        0 = 0 + (-a*e/d + b)*y + (-a*f/d + c)*z
        0 = -x + (-e/d)*y + (-f/d)*z
        0 = 0 + (h - e*g/d)*y + (i - f*g/d)*z
        """
        _, _, Mij = M[i, :], M[:, j], M[i, j]
        if Mij == 0:
            raise ZeroDivisionError("Tried to pivot about zero-valued entry.")
        A = M.copy()
        A[i, :] = -A[i, :]/Mij
        for row in range(M.shape[0]):
            if row != i:
                A[row, :] = A[row, :] + A[row, j] * A[i, :]

        return A


def _sep_const_coeff(expr):
    """
    Example
    =======

    >>> from sympy.logic.algorithms.lra_theory import _sep_const_coeff
    >>> from sympy.abc import x, y
    >>> _sep_const_coeff(2*x)
    (x, 2)
    >>> _sep_const_coeff(2*x + 3*y)
    (2*x + 3*y, 1)
    """
    if isinstance(expr, Add):
        return expr, sympify(1)

    if isinstance(expr, Mul):
        coeffs = expr.args
    else:
        coeffs = [expr]

    var, const = [], []
    for c in coeffs:
        c = sympify(c)
        if len(c.free_symbols)==0:
            const.append(c)
        else:
            var.append(c)
    return Mul(*var), Mul(*const)


def _list_terms(expr):
    if not isinstance(expr, Add):
        return [expr]

    return expr.args


def _sep_const_terms(expr):
    """
    Example
    =======

    >>> from sympy.logic.algorithms.lra_theory import _sep_const_terms
    >>> from sympy.abc import x, y
    >>> _sep_const_terms(2*x + 3*y + 2)
    (2*x + 3*y, 2)
    """
    if isinstance(expr, Add):
        terms = expr.args
    else:
        terms = [expr]

    var, const = [], []
    for t in terms:
        if len(t.free_symbols) == 0:
            const.append(t)
        else:
            var.append(t)
    return sum(var), sum(const)


def _eval_binrel(binrel):
    """
    Simplify binary relation to True / False if possible.
    """
    if not (len(binrel.lhs.free_symbols) == 0 and len(binrel.rhs.free_symbols) == 0):
        return binrel
    if binrel.function == Q.lt:
        res = binrel.lhs < binrel.rhs
    elif binrel.function == Q.gt:
        res = binrel.lhs > binrel.rhs
    elif binrel.function == Q.le:
        res = binrel.lhs <= binrel.rhs
    elif binrel.function == Q.ge:
        res = binrel.lhs >= binrel.rhs
    elif binrel.function == Q.eq:
        res = Eq(binrel.lhs, binrel.rhs)
    elif binrel.function == Q.ne:
        res = Ne(binrel.lhs, binrel.rhs)

    if res == True or res == False:
        return res
    else:
        return None


class Boundary:
    """
    Represents an upper or lower bound or an equality between a symbol
    and some constant.
    """
    def __init__(self, var, const, upper, equality, strict=None):
        if not equality in [True, False]:
            assert equality in [True, False]


        self.var = var
        if isinstance(const, tuple):
            s = const[1] != 0
            if strict:
                assert s == strict
            self.bound = const[0]
            self.strict = s
        else:
            self.bound = const
            self.strict = strict
        self.upper = upper if not equality else None
        self.equality = equality
        self.strict = strict
        assert self.strict is not None

    @staticmethod
    def from_upper(var):
        neg = -1 if var.upper_from_neg else 1
        b = Boundary(var, var.upper[0], True, var.upper_from_eq, var.upper[1] != 0)
        if neg < 0:
            b = b.get_negated()
        return b, neg

    @staticmethod
    def from_lower(var):
        neg = -1 if var.lower_from_neg else 1
        b = Boundary(var, var.lower[0], False, var.lower_from_eq, var.lower[1] != 0)
        if neg < 0:
            b = b.get_negated()
        return b, neg

    def get_negated(self):
        return Boundary(self.var, self.bound, not self.upper, self.equality, not self.strict)

    def get_inequality(self):
        if self.equality:
            return Eq(self.var.var, self.bound)
        elif self.upper and self.strict:
            return self.var.var < self.bound
        elif not self.upper and self.strict:
            return self.var.var > self.bound
        elif self.upper:
            return self.var.var <= self.bound
        else:
            return self.var.var >= self.bound

    def __repr__(self):
        return repr("Boundary(" + repr(self.get_inequality()) + ")")

    def __eq__(self, other):
        other = (other.var, other.bound, other.strict, other.upper, other.equality)
        return (self.var, self.bound, self.strict, self.upper, self.equality) == other

    def __hash__(self):
        return hash((self.var, self.bound, self.strict, self.upper, self.equality))


class LRARational():
    """
    Represents a rational plus or minus some amount
    of arbitrary small deltas.
    """
    def __init__(self, rational, delta):
        self.value = (rational, delta)

    def __lt__(self, other):
        return self.value < other.value

    def __le__(self, other):
        return self.value <= other.value

    def __eq__(self, other):
        return self.value == other.value

    def __add__(self, other):
        return LRARational(self.value[0] + other.value[0], self.value[1] + other.value[1])

    def __sub__(self, other):
        return LRARational(self.value[0] - other.value[0], self.value[1] - other.value[1])

    def __mul__(self, other):
        assert not isinstance(other, LRARational)
        return LRARational(self.value[0] * other, self.value[1] * other)

    def __getitem__(self, index):
        return self.value[index]

    def __repr__(self):
        return repr(self.value)


class LRAVariable():
    """
    Object to keep track of upper and lower bounds
    on `self.var`.
    """
    def __init__(self, var):
        self.upper = LRARational(float("inf"), 0)
        self.upper_from_eq = False
        self.upper_from_neg = False
        self.lower = LRARational(-float("inf"), 0)
        self.lower_from_eq = False
        self.lower_from_neg = False
        self.assign = LRARational(0,0)
        self.var = var
        self.col_idx = None

    def __repr__(self):
        return repr(self.var)

    def __eq__(self, other):
        if not isinstance(other, LRAVariable):
            return False
        return other.var == self.var

    def __hash__(self):
        return hash(self.var)

# === atelier-kyo-manager/buyma_research_tool.py ===
import json
from typing import List, Dict, Any

# pricing_calculator.py から利益計算関数をインポート
from pricing_calculator import calculate_profit, get_exchange_rate # get_exchange_rateもインポート

class BuymaResearchTool:
    def __init__(self):
        # 今後の機能拡張のために初期化
        pass

    def analyze_product_info(self, product_query: str) -> Dict[str, str]:
        """
        商品名または型番から商品情報を解析し、カテゴリなどを推定します。
        現時点ではシンプルなルールベースの推定を行います。
        将来的にはGPT連携を考慮します。
        """
        brand = "不明"
        category = "未分類"
        description = f"'{product_query}'に関する一般的な商品。"

        # 例: シンプルなキーワードマッチング
        if "Nike Air Max" in product_query:
            brand = "Nike"
            category = "メンズ > 靴 > スニーカー"
            description = "軽量で快適なAir Maxモデル。"
        elif "Louis Vuitton" in product_query or "LV" in product_query:
            brand = "Louis Vuitton"
            category = "レディース > バッグ・カバン > トートバッグ" # 例
            description = "高級感あふれるルイ・ヴィトンのアイテム。"
        # 他のルールを追加...

        return {
            "brand": brand,
            "category": category,
            "description": description
        }

    def search_suppliers(self, product_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        海外ECサイトから仕入れ先情報を検索します。
        この部分はWebスクレイピングやAPI連携が必要なため、現在はモックデータです。
        """
        print(f"DEBUG: Searching suppliers for {product_info['brand']} {product_info['description']}")
        # 実際のスクレイピング/API呼び出しの代わりにモックデータを返します
        mock_suppliers = [
            {
                "source_name": "SSENSE",
                "source_url": "https://www.ssense.com/mock_product_url",
                "source_price": 120, # USD
                "shipping_cost": 20, # USD
                "currency": "USD",
                "in_stock": True,
                "variants": {"27cm": "在庫あり", "28cm": "在庫切れ"}
            },
            {
                "source_name": "Farfetch",
                "source_url": "https://www.farfetch.com/mock_product_url",
                "source_price": 130, # USD
                "shipping_cost": 15, # USD
                "currency": "USD",
                "in_stock": False,
                "variants": {"27cm": "在庫切れ", "28cm": "在庫切れ"}
            }
        ]
        return mock_suppliers

    def get_buyma_sales_status(self, product_info: Dict[str, Any]) -> Dict[str, Any]:
        """
        BUYMAでの販売状況を取得します。
        この部分はSeleniumなどによるクロールが必要なため、現在はモックデータです。
        """
        print(f"DEBUG: Getting BUYMA sales status for {product_info['brand']} {product_info['description']}")
        # 実際のクロールの代わりにモックデータを返します
        return {
            "buyma_lowest_price": 25300,
            "buyma_competitors": 7,
            "buyma_sales_recent": "30日以内に2件販売"
        }

    def calculate_and_rank_profit(
        self,
        buyma_target_price: float,
        suppliers: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        各仕入れ先からの利益を計算し、利益率に基づいてランキングします。
        """
        profitable_options = []
        for supplier in suppliers:
            if supplier["in_stock"]:
                profit_data = calculate_profit(
                    buyma_price=buyma_target_price,
                    source_price=supplier["source_price"],
                    shipping_cost=supplier["shipping_cost"],
                    source_currency=supplier["currency"] # 仕入れ元の通貨を渡す
                )
                profit_estimate = profit_data["profit_estimate"]
                profit_rate = profit_data["profit_rate"]

                if profit_rate >= 10: # 利益率10%以上を候補とする
                    supplier_with_profit = supplier.copy()
                    supplier_with_profit["profit_estimate"] = profit_estimate
                    supplier_with_profit["profit_rate"] = f"{profit_rate}%"
                    supplier_with_profit["is_profitable"] = True
                    # 優先度スコアは仮で利益率をベースに設定
                    supplier_with_profit["priority_score"] = int(profit_rate)
                    profitable_options.append(supplier_with_profit)
            else:
                # 在庫がない場合は利益計算対象外
                supplier_with_profit = supplier.copy()
                supplier_with_profit["profit_estimate"] = 0
                supplier_with_profit["profit_rate"] = "0%"
                supplier_with_profit["is_profitable"] = False
                supplier_with_profit["priority_score"] = 0
                profitable_options.append(supplier_with_profit)


        # 利益率の高い順にソート
        profitable_options.sort(key=lambda x: float(x.get("profit_rate", "0%").replace('%', '')), reverse=True)
        return profitable_options

    def run_research(self, product_query: str) -> List[Dict[str, Any]]:
        """
        一連のリサーチプロセスを実行し、結果を返します。
        """
        print(f"STEP 1: Analyzing product information for '{product_query}'...")
        product_info = self.analyze_product_info(product_query)
        print(f"  - Brand: {product_info['brand']}, Category: {product_info['category']}")

        print("STEP 2: Searching for suppliers on cross-border EC sites...")
        suppliers = self.search_suppliers(product_info)
        print(f"  - Found {len(suppliers)} potential suppliers (mock data).")

        print("STEP 3: Getting BUYMA sales status...")
        buyma_status = self.get_buyma_sales_status(product_info)
        print(f"  - BUYMA Lowest Price: {buyma_status['buyma_lowest_price']}円, Competitors: {buyma_status['buyma_competitors']}")

        # BUYMAの目標販売価格を仮で設定（最低販売価格を参考に）
        buyma_target_price = buyma_status["buyma_lowest_price"]

        print("STEP 4: Calculating profit and ranking options...")
        ranked_options = self.calculate_and_rank_profit(buyma_target_price, suppliers)

        final_results = []
        for option in ranked_options:
            # source_total_price を動的に計算
            exchange_rate_to_jpy = get_exchange_rate(option["currency"], "JPY")
            source_total_price_jpy = (option["source_price"] + option["shipping_cost"]) * exchange_rate_to_jpy

            result = {
                "title": f"{product_info['brand']} {product_info['description']}",
                "brand": product_info["brand"],
                "category": product_info["category"],
                "buyma_price": buyma_target_price,
                "description": product_info["description"],
                "variants": option.get("variants", {}),
                "source_url": option["source_url"],
                "source_name": option["source_name"],
                "source_total_price": round(source_total_price_jpy, 2), # 動的に計算された合計仕入れ価格
                "profit_estimate": option["profit_estimate"],
                "profit_rate": option["profit_rate"],
                "buyma_competitors": buyma_status["buyma_competitors"],
                "buyma_lowest_price": buyma_status["buyma_lowest_price"],
                "buyma_sales_recent": buyma_status["buyma_sales_recent"],
                "is_profitable": option["is_profitable"],
                "priority_score": option["priority_score"]
            }
            final_results.append(result)

        print("Research complete. Generating output.")
        return final_results

# テスト用の使用例
if __name__ == "__main__":
    tool = BuymaResearchTool()
    query = "Nike Air Max 97 White Metallic Silver"
    results = tool.run_research(query)

    print("\n--- Research Results ---")
    print(json.dumps(results, indent=2, ensure_ascii=False))

    query_lv = "Louis Vuitton Neverfull MM"
    results_lv = tool.run_research(query_lv)

    print("\n--- Research Results (Louis Vuitton) ---")
    print(json.dumps(results_lv, indent=2, ensure_ascii=False))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\worksheet\worksheet.py ===
# Copyright (c) 2010-2024 openpyxl

"""Worksheet is the 2nd-level container in Excel."""


# Python stdlib imports
from itertools import chain
from operator import itemgetter
from inspect import isgenerator
from warnings import warn

# compatibility imports
from openpyxl.compat import (
    deprecated,
)

# package imports
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
    coordinate_to_tuple,
)
from openpyxl.cell import Cell, MergedCell
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.packaging.relationship import RelationshipList
from openpyxl.workbook.child import _WorkbookChild
from openpyxl.workbook.defined_name import (
    DefinedNameDict,
)

from openpyxl.formula.translate import Translator

from .datavalidation import DataValidationList
from .page import (
    PrintPageSetup,
    PageMargins,
    PrintOptions,
)
from .dimensions import (
    ColumnDimension,
    RowDimension,
    DimensionHolder,
    SheetFormatProperties,
)
from .protection import SheetProtection
from .filters import AutoFilter
from .views import (
    Pane,
    Selection,
    SheetViewList,
)
from .cell_range import MultiCellRange, CellRange
from .merge import MergedCellRange
from .properties import WorksheetProperties
from .pagebreak import RowBreak, ColBreak
from .scenario import ScenarioList
from .table import TableList
from .formula import ArrayFormula
from .print_settings import (
    PrintTitles,
    ColRange,
    RowRange,
    PrintArea,
)


class Worksheet(_WorkbookChild):
    """Represents a worksheet.

    Do not create worksheets yourself,
    use :func:`openpyxl.workbook.Workbook.create_sheet` instead

    """

    _rel_type = "worksheet"
    _path = "/xl/worksheets/sheet{0}.xml"
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"

    BREAK_NONE = 0
    BREAK_ROW = 1
    BREAK_COLUMN = 2

    SHEETSTATE_VISIBLE = 'visible'
    SHEETSTATE_HIDDEN = 'hidden'
    SHEETSTATE_VERYHIDDEN = 'veryHidden'

    # Paper size
    PAPERSIZE_LETTER = '1'
    PAPERSIZE_LETTER_SMALL = '2'
    PAPERSIZE_TABLOID = '3'
    PAPERSIZE_LEDGER = '4'
    PAPERSIZE_LEGAL = '5'
    PAPERSIZE_STATEMENT = '6'
    PAPERSIZE_EXECUTIVE = '7'
    PAPERSIZE_A3 = '8'
    PAPERSIZE_A4 = '9'
    PAPERSIZE_A4_SMALL = '10'
    PAPERSIZE_A5 = '11'

    # Page orientation
    ORIENTATION_PORTRAIT = 'portrait'
    ORIENTATION_LANDSCAPE = 'landscape'

    def __init__(self, parent, title=None):
        _WorkbookChild.__init__(self, parent, title)
        self._setup()

    def _setup(self):
        self.row_dimensions = DimensionHolder(worksheet=self,
                                              default_factory=self._add_row)
        self.column_dimensions = DimensionHolder(worksheet=self,
                                                 default_factory=self._add_column)
        self.row_breaks = RowBreak()
        self.col_breaks = ColBreak()
        self._cells = {}
        self._charts = []
        self._images = []
        self._rels = RelationshipList()
        self._drawing = None
        self._comments = []
        self.merged_cells = MultiCellRange()
        self._tables = TableList()
        self._pivots = []
        self.data_validations = DataValidationList()
        self._hyperlinks = []
        self.sheet_state = 'visible'
        self.page_setup = PrintPageSetup(worksheet=self)
        self.print_options = PrintOptions()
        self._print_rows = None
        self._print_cols = None
        self._print_area = PrintArea()
        self.page_margins = PageMargins()
        self.views = SheetViewList()
        self.protection = SheetProtection()
        self.defined_names = DefinedNameDict()

        self._current_row = 0
        self.auto_filter = AutoFilter()
        self.conditional_formatting = ConditionalFormattingList()
        self.legacy_drawing = None
        self.sheet_properties = WorksheetProperties()
        self.sheet_format = SheetFormatProperties()
        self.scenarios = ScenarioList()


    @property
    def sheet_view(self):
        return self.views.active


    @property
    def selected_cell(self):
        return self.sheet_view.selection[0].sqref


    @property
    def active_cell(self):
        return self.sheet_view.selection[0].activeCell


    @property
    def array_formulae(self):
        """Returns a dictionary of cells with array formulae and the cells in array"""
        result = {}
        for c in self._cells.values():
            if c.data_type == "f":
                if isinstance(c.value, ArrayFormula):
                    result[c.coordinate] = c.value.ref
        return result


    @property
    def show_gridlines(self):
        return self.sheet_view.showGridLines


    @property
    def freeze_panes(self):
        if self.sheet_view.pane is not None:
            return self.sheet_view.pane.topLeftCell


    @freeze_panes.setter
    def freeze_panes(self, topLeftCell=None):
        if isinstance(topLeftCell, Cell):
            topLeftCell = topLeftCell.coordinate
        if topLeftCell == 'A1':
            topLeftCell = None

        if not topLeftCell:
            self.sheet_view.pane = None
            return

        row, column = coordinate_to_tuple(topLeftCell)

        view = self.sheet_view
        view.pane = Pane(topLeftCell=topLeftCell,
                        activePane="topRight",
                        state="frozen")
        view.selection[0].pane = "topRight"

        if column > 1:
            view.pane.xSplit = column - 1
        if row > 1:
            view.pane.ySplit = row - 1
            view.pane.activePane = 'bottomLeft'
            view.selection[0].pane = "bottomLeft"
            if column > 1:
                view.selection[0].pane = "bottomRight"
                view.pane.activePane = 'bottomRight'

        if row > 1 and column > 1:
            sel = list(view.selection)
            sel.insert(0, Selection(pane="topRight", activeCell=None, sqref=None))
            sel.insert(1, Selection(pane="bottomLeft", activeCell=None, sqref=None))
            view.selection = sel


    def cell(self, row, column, value=None):
        """
        Returns a cell object based on the given coordinates.

        Usage: cell(row=15, column=1, value=5)

        Calling `cell` creates cells in memory when they
        are first accessed.

        :param row: row index of the cell (e.g. 4)
        :type row: int

        :param column: column index of the cell (e.g. 3)
        :type column: int

        :param value: value of the cell (e.g. 5)
        :type value: numeric or time or string or bool or none

        :rtype: openpyxl.cell.cell.Cell
        """

        if row < 1 or column < 1:
            raise ValueError("Row or column values must be at least 1")

        cell = self._get_cell(row, column)
        if value is not None:
            cell.value = value

        return cell


    def _get_cell(self, row, column):
        """
        Internal method for getting a cell from a worksheet.
        Will create a new cell if one doesn't already exist.
        """
        if not 0 < row < 1048577:
            raise ValueError(f"Row numbers must be between 1 and 1048576. Row number supplied was {row}")
        coordinate = (row, column)
        if not coordinate in self._cells:
            cell = Cell(self, row=row, column=column)
            self._add_cell(cell)
        return self._cells[coordinate]


    def _add_cell(self, cell):
        """
        Internal method for adding cell objects.
        """
        column = cell.col_idx
        row = cell.row
        self._current_row = max(row, self._current_row)
        self._cells[(row, column)] = cell


    def __getitem__(self, key):
        """Convenience access by Excel style coordinates

        The key can be a single cell coordinate 'A1', a range of cells 'A1:D25',
        individual rows or columns 'A', 4 or ranges of rows or columns 'A:D',
        4:10.

        Single cells will always be created if they do not exist.

        Returns either a single cell or a tuple of rows or columns.
        """
        if isinstance(key, slice):
            if not all([key.start, key.stop]):
                raise IndexError("{0} is not a valid coordinate or range".format(key))
            key = "{0}:{1}".format(key.start, key.stop)

        if isinstance(key, int):
            key = str(key
                      )
        min_col, min_row, max_col, max_row = range_boundaries(key)

        if not any([min_col, min_row, max_col, max_row]):
            raise IndexError("{0} is not a valid coordinate or range".format(key))

        if min_row is None:
            cols = tuple(self.iter_cols(min_col, max_col))
            if min_col == max_col:
                cols = cols[0]
            return cols
        if min_col is None:
            rows = tuple(self.iter_rows(min_col=min_col, min_row=min_row,
                                        max_col=self.max_column, max_row=max_row))
            if min_row == max_row:
                rows = rows[0]
            return rows
        if ":" not in key:
            return self._get_cell(min_row, min_col)
        return tuple(self.iter_rows(min_row=min_row, min_col=min_col,
                                    max_row=max_row, max_col=max_col))


    def __setitem__(self, key, value):
        self[key].value = value


    def __iter__(self):
        return self.iter_rows()


    def __delitem__(self, key):
        row, column = coordinate_to_tuple(key)
        if (row, column) in self._cells:
            del self._cells[(row, column)]


    @property
    def min_row(self):
        """The minimum row index containing data (1-based)

        :type: int
        """
        min_row = 1
        if self._cells:
            min_row = min(self._cells)[0]
        return min_row


    @property
    def max_row(self):
        """The maximum row index containing data (1-based)

        :type: int
        """
        max_row = 1
        if self._cells:
            max_row = max(self._cells)[0]
        return max_row


    @property
    def min_column(self):
        """The minimum column index containing data (1-based)

        :type: int
        """
        min_col = 1
        if self._cells:
            min_col = min(c[1] for c in self._cells)
        return min_col


    @property
    def max_column(self):
        """The maximum column index containing data (1-based)

        :type: int
        """
        max_col = 1
        if self._cells:
            max_col = max(c[1] for c in self._cells)
        return max_col


    def calculate_dimension(self):
        """Return the minimum bounding range for all cells containing data (ex. 'A1:M24')

        :rtype: string
        """
        if self._cells:
            rows = set()
            cols = set()
            for row, col in self._cells:
                rows.add(row)
                cols.add(col)
            max_row = max(rows)
            max_col = max(cols)
            min_col = min(cols)
            min_row = min(rows)
        else:
            return "A1:A1"

        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


    @property
    def dimensions(self):
        """Returns the result of :func:`calculate_dimension`"""
        return self.calculate_dimension()


    def iter_rows(self, min_row=None, max_row=None, min_col=None, max_col=None, values_only=False):
        """
        Produces cells from the worksheet, by row. Specify the iteration range
        using indices of rows and columns.

        If no indices are specified the range starts at A1.

        If no cells are in the worksheet an empty tuple will be returned.

        :param min_col: smallest column index (1-based index)
        :type min_col: int

        :param min_row: smallest row index (1-based index)
        :type min_row: int

        :param max_col: largest column index (1-based index)
        :type max_col: int

        :param max_row: largest row index (1-based index)
        :type max_row: int

        :param values_only: whether only cell values should be returned
        :type values_only: bool

        :rtype: generator
        """

        if self._current_row == 0 and not any([min_col, min_row, max_col, max_row ]):
            return iter(())


        min_col = min_col or 1
        min_row = min_row or 1
        max_col = max_col or self.max_column
        max_row = max_row or self.max_row

        return self._cells_by_row(min_col, min_row, max_col, max_row, values_only)


    def _cells_by_row(self, min_col, min_row, max_col, max_row, values_only=False):
        for row in range(min_row, max_row + 1):
            cells = (self.cell(row=row, column=column) for column in range(min_col, max_col + 1))
            if values_only:
                yield tuple(cell.value for cell in cells)
            else:
                yield tuple(cells)


    @property
    def rows(self):
        """Produces all cells in the worksheet, by row (see :func:`iter_rows`)

        :type: generator
        """
        return self.iter_rows()


    @property
    def values(self):
        """Produces all cell values in the worksheet, by row

        :type: generator
        """
        for row in self.iter_rows(values_only=True):
            yield row


    def iter_cols(self, min_col=None, max_col=None, min_row=None, max_row=None, values_only=False):
        """
        Produces cells from the worksheet, by column. Specify the iteration range
        using indices of rows and columns.

        If no indices are specified the range starts at A1.

        If no cells are in the worksheet an empty tuple will be returned.

        :param min_col: smallest column index (1-based index)
        :type min_col: int

        :param min_row: smallest row index (1-based index)
        :type min_row: int

        :param max_col: largest column index (1-based index)
        :type max_col: int

        :param max_row: largest row index (1-based index)
        :type max_row: int

        :param values_only: whether only cell values should be returned
        :type values_only: bool

        :rtype: generator
        """

        if self._current_row == 0 and not any([min_col, min_row, max_col, max_row]):
            return iter(())

        min_col = min_col or 1
        min_row = min_row or 1
        max_col = max_col or self.max_column
        max_row = max_row or self.max_row

        return self._cells_by_col(min_col, min_row, max_col, max_row, values_only)


    def _cells_by_col(self, min_col, min_row, max_col, max_row, values_only=False):
        """
        Get cells by column
        """
        for column in range(min_col, max_col+1):
            cells = (self.cell(row=row, column=column)
                        for row in range(min_row, max_row+1))
            if values_only:
                yield tuple(cell.value for cell in cells)
            else:
                yield tuple(cells)


    @property
    def columns(self):
        """Produces all cells in the worksheet, by column  (see :func:`iter_cols`)"""
        return self.iter_cols()


    @property
    def column_groups(self):
        """
        Return a list of column ranges where more than one column
        """
        return [cd.range for cd in self.column_dimensions.values() if cd.min and cd.max > cd.min]


    def set_printer_settings(self, paper_size, orientation):
        """Set printer settings """

        self.page_setup.paperSize = paper_size
        self.page_setup.orientation = orientation


    def add_data_validation(self, data_validation):
        """ Add a data-validation object to the sheet.  The data-validation
            object defines the type of data-validation to be applied and the
            cell or range of cells it should apply to.
        """
        self.data_validations.append(data_validation)


    def add_chart(self, chart, anchor=None):
        """
        Add a chart to the sheet
        Optionally provide a cell for the top-left anchor
        """
        if anchor is not None:
            chart.anchor = anchor
        self._charts.append(chart)


    def add_image(self, img, anchor=None):
        """
        Add an image to the sheet.
        Optionally provide a cell for the top-left anchor
        """
        if anchor is not None:
            img.anchor = anchor
        self._images.append(img)


    def add_table(self, table):
        """
        Check for duplicate name in definedNames and other worksheet tables
        before adding table.
        """

        if self.parent._duplicate_name(table.name):
            raise ValueError("Table with name {0} already exists".format(table.name))
        if not hasattr(self, "_get_cell"):
            warn("In write-only mode you must add table columns manually")
        self._tables.add(table)


    @property
    def tables(self):
        return self._tables


    def add_pivot(self, pivot):
        self._pivots.append(pivot)


    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1) """
        if range_string is None:
            cr = CellRange(range_string=range_string, min_col=start_column, min_row=start_row,
                      max_col=end_column, max_row=end_row)
            range_string = cr.coord
        mcr = MergedCellRange(self, range_string)
        self.merged_cells.add(mcr)
        self._clean_merge_range(mcr)


    def _clean_merge_range(self, mcr):
        """
        Remove all but the top left-cell from a range of merged cells
        and recreate the lost border information.
        Borders are then applied
        """
        cells = mcr.cells
        next(cells) # skip first cell
        for row, col in cells:
            self._cells[row, col] = MergedCell(self, row, col)
        mcr.format()


    @property
    @deprecated("Use ws.merged_cells.ranges")
    def merged_cell_ranges(self):
        """Return a copy of cell ranges"""
        return self.merged_cells.ranges[:]


    def unmerge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Remove merge on a cell range.  Range is a cell range (e.g. A1:E1) """
        cr = CellRange(range_string=range_string, min_col=start_column, min_row=start_row,
                      max_col=end_column, max_row=end_row)

        if cr.coord not in self.merged_cells:
            raise ValueError("Cell range {0} is not merged".format(cr.coord))

        self.merged_cells.remove(cr)

        cells = cr.cells
        next(cells) # skip first cell
        for row, col in cells:
            del self._cells[(row, col)]


    def append(self, iterable):
        """Appends a group of values at the bottom of the current sheet.

        * If it's a list: all values are added in order, starting from the first column
        * If it's a dict: values are assigned to the columns indicated by the keys (numbers or letters)

        :param iterable: list, range or generator, or dict containing values to append
        :type iterable: list|tuple|range|generator or dict

        Usage:

        * append(['This is A1', 'This is B1', 'This is C1'])
        * **or** append({'A' : 'This is A1', 'C' : 'This is C1'})
        * **or** append({1 : 'This is A1', 3 : 'This is C1'})

        :raise: TypeError when iterable is neither a list/tuple nor a dict

        """
        row_idx = self._current_row + 1

        if (isinstance(iterable, (list, tuple, range))
            or isgenerator(iterable)):
            for col_idx, content in enumerate(iterable, 1):
                if isinstance(content, Cell):
                    # compatible with write-only mode
                    cell = content
                    if cell.parent and cell.parent != self:
                        raise ValueError("Cells cannot be copied from other worksheets")
                    cell.parent = self
                    cell.column = col_idx
                    cell.row = row_idx
                else:
                    cell = Cell(self, row=row_idx, column=col_idx, value=content)
                self._cells[(row_idx, col_idx)] = cell

        elif isinstance(iterable, dict):
            for col_idx, content in iterable.items():
                if isinstance(col_idx, str):
                    col_idx = column_index_from_string(col_idx)
                cell = Cell(self, row=row_idx, column=col_idx, value=content)
                self._cells[(row_idx, col_idx)] = cell

        else:
            self._invalid_row(iterable)

        self._current_row = row_idx


    def _move_cells(self, min_row=None, min_col=None, offset=0, row_or_col="row"):
        """
        Move either rows or columns around by the offset
        """
        reverse = offset > 0 # start at the end if inserting
        row_offset = 0
        col_offset = 0

        # need to make affected ranges contiguous
        if row_or_col == 'row':
            cells = self.iter_rows(min_row=min_row)
            row_offset = offset
            key = 0
        else:
            cells = self.iter_cols(min_col=min_col)
            col_offset = offset
            key = 1
        cells = list(cells)

        for row, column in sorted(self._cells, key=itemgetter(key), reverse=reverse):
            if min_row and row < min_row:
                continue
            elif min_col and column < min_col:
                continue

            self._move_cell(row, column, row_offset, col_offset)


    def insert_rows(self, idx, amount=1):
        """
        Insert row or rows before row==idx
        """
        self._move_cells(min_row=idx, offset=amount, row_or_col="row")
        self._current_row = self.max_row


    def insert_cols(self, idx, amount=1):
        """
        Insert column or columns before col==idx
        """
        self._move_cells(min_col=idx, offset=amount, row_or_col="column")


    def delete_rows(self, idx, amount=1):
        """
        Delete row or rows from row==idx
        """

        remainder = _gutter(idx, amount, self.max_row)

        self._move_cells(min_row=idx+amount, offset=-amount, row_or_col="row")

        # calculating min and max col is an expensive operation, do it only once
        min_col = self.min_column
        max_col = self.max_column + 1
        for row in remainder:
            for col in range(min_col, max_col):
                if (row, col) in self._cells:
                    del self._cells[row, col]
        self._current_row = self.max_row
        if not self._cells:
            self._current_row = 0


    def delete_cols(self, idx, amount=1):
        """
        Delete column or columns from col==idx
        """

        remainder = _gutter(idx, amount, self.max_column)

        self._move_cells(min_col=idx+amount, offset=-amount, row_or_col="column")

        # calculating min and max row is an expensive operation, do it only once
        min_row = self.min_row
        max_row = self.max_row + 1
        for col in remainder:
            for row in range(min_row, max_row):
                if (row, col) in self._cells:
                    del self._cells[row, col]


    def move_range(self, cell_range, rows=0, cols=0, translate=False):
        """
        Move a cell range by the number of rows and/or columns:
        down if rows > 0 and up if rows < 0
        right if cols > 0 and left if cols < 0
        Existing cells will be overwritten.
        Formulae and references will not be updated.
        """
        if isinstance(cell_range, str):
            cell_range = CellRange(cell_range)
        if not isinstance(cell_range, CellRange):
            raise ValueError("Only CellRange objects can be moved")
        if not rows and not cols:
            return

        down = rows > 0
        right = cols > 0

        if rows:
            cells = sorted(cell_range.rows, reverse=down)
        else:
            cells = sorted(cell_range.cols, reverse=right)

        for row, col in chain.from_iterable(cells):
            self._move_cell(row, col, rows, cols, translate)

        # rebase moved range
        cell_range.shift(row_shift=rows, col_shift=cols)


    def _move_cell(self, row, column, row_offset, col_offset, translate=False):
        """
        Move a cell from one place to another.
        Delete at old index
        Rebase coordinate
        """
        cell = self._get_cell(row, column)
        new_row = cell.row + row_offset
        new_col = cell.column + col_offset
        self._cells[new_row, new_col] = cell
        del self._cells[(cell.row, cell.column)]
        cell.row = new_row
        cell.column = new_col
        if translate and cell.data_type == "f":
            t = Translator(cell.value, cell.coordinate)
            cell.value = t.translate_formula(row_delta=row_offset, col_delta=col_offset)


    def _invalid_row(self, iterable):
        raise TypeError('Value must be a list, tuple, range or generator, or a dict. Supplied value is {0}'.format(
            type(iterable))
                        )


    def _add_column(self):
        """Dimension factory for column information"""

        return ColumnDimension(self)

    def _add_row(self):
        """Dimension factory for row information"""

        return RowDimension(self)


    @property
    def print_title_rows(self):
        """Rows to be printed at the top of every page (ex: '1:3')"""
        if self._print_rows:
            return str(self._print_rows)


    @print_title_rows.setter
    def print_title_rows(self, rows):
        """
        Set rows to be printed on the top of every page
        format `1:3`
        """
        if rows is not None:
            self._print_rows = RowRange(rows)


    @property
    def print_title_cols(self):
        """Columns to be printed at the left side of every page (ex: 'A:C')"""
        if self._print_cols:
            return str(self._print_cols)


    @print_title_cols.setter
    def print_title_cols(self, cols):
        """
        Set cols to be printed on the left of every page
        format ``A:C`
        """
        if cols is not None:
            self._print_cols = ColRange(cols)


    @property
    def print_titles(self):
        titles = PrintTitles(cols=self._print_cols, rows=self._print_rows, title=self.title)
        return str(titles)


    @property
    def print_area(self):
        """
        The print area for the worksheet, or None if not set. To set, supply a range
        like 'A1:D4' or a list of ranges.
        """
        self._print_area.title = self.title
        return str(self._print_area)


    @print_area.setter
    def print_area(self, value):
        """
        Range of cells in the form A1:D4 or list of ranges. Print area can be cleared
        by passing `None` or an empty list
        """
        if not value:
            self._print_area = PrintArea()
        elif isinstance(value, str):
            self._print_area = PrintArea.from_string(value)
        elif hasattr(value, "__iter__"):
            self._print_area = PrintArea.from_string(",".join(value))


def _gutter(idx, offset, max_val):
    """
    When deleting rows and columns are deleted we rely on overwriting.
    This may not be the case for a large offset on small set of cells:
    range(cells_to_delete) > range(cell_to_be_moved)
    """
    gutter = range(max(max_val+1-offset, idx), min(idx+offset, max_val)+1)
    return gutter

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pycparser\ply\cpp.py ===
# -----------------------------------------------------------------------------
# cpp.py
#
# Author:  David Beazley (http://www.dabeaz.com)
# Copyright (C) 2017
# All rights reserved
#
# This module implements an ANSI-C style lexical preprocessor for PLY.
# -----------------------------------------------------------------------------
import sys

# Some Python 3 compatibility shims
if sys.version_info.major < 3:
    STRING_TYPES = (str, unicode)
else:
    STRING_TYPES = str
    xrange = range

# -----------------------------------------------------------------------------
# Default preprocessor lexer definitions.   These tokens are enough to get
# a basic preprocessor working.   Other modules may import these if they want
# -----------------------------------------------------------------------------

tokens = (
   'CPP_ID','CPP_INTEGER', 'CPP_FLOAT', 'CPP_STRING', 'CPP_CHAR', 'CPP_WS', 'CPP_COMMENT1', 'CPP_COMMENT2', 'CPP_POUND','CPP_DPOUND'
)

literals = "+-*/%|&~^<>=!?()[]{}.,;:\\\'\""

# Whitespace
def t_CPP_WS(t):
    r'\s+'
    t.lexer.lineno += t.value.count("\n")
    return t

t_CPP_POUND = r'\#'
t_CPP_DPOUND = r'\#\#'

# Identifier
t_CPP_ID = r'[A-Za-z_][\w_]*'

# Integer literal
def CPP_INTEGER(t):
    r'(((((0x)|(0X))[0-9a-fA-F]+)|(\d+))([uU][lL]|[lL][uU]|[uU]|[lL])?)'
    return t

t_CPP_INTEGER = CPP_INTEGER

# Floating literal
t_CPP_FLOAT = r'((\d+)(\.\d+)(e(\+|-)?(\d+))? | (\d+)e(\+|-)?(\d+))([lL]|[fF])?'

# String literal
def t_CPP_STRING(t):
    r'\"([^\\\n]|(\\(.|\n)))*?\"'
    t.lexer.lineno += t.value.count("\n")
    return t

# Character constant 'c' or L'c'
def t_CPP_CHAR(t):
    r'(L)?\'([^\\\n]|(\\(.|\n)))*?\''
    t.lexer.lineno += t.value.count("\n")
    return t

# Comment
def t_CPP_COMMENT1(t):
    r'(/\*(.|\n)*?\*/)'
    ncr = t.value.count("\n")
    t.lexer.lineno += ncr
    # replace with one space or a number of '\n'
    t.type = 'CPP_WS'; t.value = '\n' * ncr if ncr else ' '
    return t

# Line comment
def t_CPP_COMMENT2(t):
    r'(//.*?(\n|$))'
    # replace with '/n'
    t.type = 'CPP_WS'; t.value = '\n'
    return t

def t_error(t):
    t.type = t.value[0]
    t.value = t.value[0]
    t.lexer.skip(1)
    return t

import re
import copy
import time
import os.path

# -----------------------------------------------------------------------------
# trigraph()
#
# Given an input string, this function replaces all trigraph sequences.
# The following mapping is used:
#
#     ??=    #
#     ??/    \
#     ??'    ^
#     ??(    [
#     ??)    ]
#     ??!    |
#     ??<    {
#     ??>    }
#     ??-    ~
# -----------------------------------------------------------------------------

_trigraph_pat = re.compile(r'''\?\?[=/\'\(\)\!<>\-]''')
_trigraph_rep = {
    '=':'#',
    '/':'\\',
    "'":'^',
    '(':'[',
    ')':']',
    '!':'|',
    '<':'{',
    '>':'}',
    '-':'~'
}

def trigraph(input):
    return _trigraph_pat.sub(lambda g: _trigraph_rep[g.group()[-1]],input)

# ------------------------------------------------------------------
# Macro object
#
# This object holds information about preprocessor macros
#
#    .name      - Macro name (string)
#    .value     - Macro value (a list of tokens)
#    .arglist   - List of argument names
#    .variadic  - Boolean indicating whether or not variadic macro
#    .vararg    - Name of the variadic parameter
#
# When a macro is created, the macro replacement token sequence is
# pre-scanned and used to create patch lists that are later used
# during macro expansion
# ------------------------------------------------------------------

class Macro(object):
    def __init__(self,name,value,arglist=None,variadic=False):
        self.name = name
        self.value = value
        self.arglist = arglist
        self.variadic = variadic
        if variadic:
            self.vararg = arglist[-1]
        self.source = None

# ------------------------------------------------------------------
# Preprocessor object
#
# Object representing a preprocessor.  Contains macro definitions,
# include directories, and other information
# ------------------------------------------------------------------

class Preprocessor(object):
    def __init__(self,lexer=None):
        if lexer is None:
            lexer = lex.lexer
        self.lexer = lexer
        self.macros = { }
        self.path = []
        self.temp_path = []

        # Probe the lexer for selected tokens
        self.lexprobe()

        tm = time.localtime()
        self.define("__DATE__ \"%s\"" % time.strftime("%b %d %Y",tm))
        self.define("__TIME__ \"%s\"" % time.strftime("%H:%M:%S",tm))
        self.parser = None

    # -----------------------------------------------------------------------------
    # tokenize()
    #
    # Utility function. Given a string of text, tokenize into a list of tokens
    # -----------------------------------------------------------------------------

    def tokenize(self,text):
        tokens = []
        self.lexer.input(text)
        while True:
            tok = self.lexer.token()
            if not tok: break
            tokens.append(tok)
        return tokens

    # ---------------------------------------------------------------------
    # error()
    #
    # Report a preprocessor error/warning of some kind
    # ----------------------------------------------------------------------

    def error(self,file,line,msg):
        print("%s:%d %s" % (file,line,msg))

    # ----------------------------------------------------------------------
    # lexprobe()
    #
    # This method probes the preprocessor lexer object to discover
    # the token types of symbols that are important to the preprocessor.
    # If this works right, the preprocessor will simply "work"
    # with any suitable lexer regardless of how tokens have been named.
    # ----------------------------------------------------------------------

    def lexprobe(self):

        # Determine the token type for identifiers
        self.lexer.input("identifier")
        tok = self.lexer.token()
        if not tok or tok.value != "identifier":
            print("Couldn't determine identifier type")
        else:
            self.t_ID = tok.type

        # Determine the token type for integers
        self.lexer.input("12345")
        tok = self.lexer.token()
        if not tok or int(tok.value) != 12345:
            print("Couldn't determine integer type")
        else:
            self.t_INTEGER = tok.type
            self.t_INTEGER_TYPE = type(tok.value)

        # Determine the token type for strings enclosed in double quotes
        self.lexer.input("\"filename\"")
        tok = self.lexer.token()
        if not tok or tok.value != "\"filename\"":
            print("Couldn't determine string type")
        else:
            self.t_STRING = tok.type

        # Determine the token type for whitespace--if any
        self.lexer.input("  ")
        tok = self.lexer.token()
        if not tok or tok.value != "  ":
            self.t_SPACE = None
        else:
            self.t_SPACE = tok.type

        # Determine the token type for newlines
        self.lexer.input("\n")
        tok = self.lexer.token()
        if not tok or tok.value != "\n":
            self.t_NEWLINE = None
            print("Couldn't determine token for newlines")
        else:
            self.t_NEWLINE = tok.type

        self.t_WS = (self.t_SPACE, self.t_NEWLINE)

        # Check for other characters used by the preprocessor
        chars = [ '<','>','#','##','\\','(',')',',','.']
        for c in chars:
            self.lexer.input(c)
            tok = self.lexer.token()
            if not tok or tok.value != c:
                print("Unable to lex '%s' required for preprocessor" % c)

    # ----------------------------------------------------------------------
    # add_path()
    #
    # Adds a search path to the preprocessor.
    # ----------------------------------------------------------------------

    def add_path(self,path):
        self.path.append(path)

    # ----------------------------------------------------------------------
    # group_lines()
    #
    # Given an input string, this function splits it into lines.  Trailing whitespace
    # is removed.   Any line ending with \ is grouped with the next line.  This
    # function forms the lowest level of the preprocessor---grouping into text into
    # a line-by-line format.
    # ----------------------------------------------------------------------

    def group_lines(self,input):
        lex = self.lexer.clone()
        lines = [x.rstrip() for x in input.splitlines()]
        for i in xrange(len(lines)):
            j = i+1
            while lines[i].endswith('\\') and (j < len(lines)):
                lines[i] = lines[i][:-1]+lines[j]
                lines[j] = ""
                j += 1

        input = "\n".join(lines)
        lex.input(input)
        lex.lineno = 1

        current_line = []
        while True:
            tok = lex.token()
            if not tok:
                break
            current_line.append(tok)
            if tok.type in self.t_WS and '\n' in tok.value:
                yield current_line
                current_line = []

        if current_line:
            yield current_line

    # ----------------------------------------------------------------------
    # tokenstrip()
    #
    # Remove leading/trailing whitespace tokens from a token list
    # ----------------------------------------------------------------------

    def tokenstrip(self,tokens):
        i = 0
        while i < len(tokens) and tokens[i].type in self.t_WS:
            i += 1
        del tokens[:i]
        i = len(tokens)-1
        while i >= 0 and tokens[i].type in self.t_WS:
            i -= 1
        del tokens[i+1:]
        return tokens


    # ----------------------------------------------------------------------
    # collect_args()
    #
    # Collects comma separated arguments from a list of tokens.   The arguments
    # must be enclosed in parenthesis.  Returns a tuple (tokencount,args,positions)
    # where tokencount is the number of tokens consumed, args is a list of arguments,
    # and positions is a list of integers containing the starting index of each
    # argument.  Each argument is represented by a list of tokens.
    #
    # When collecting arguments, leading and trailing whitespace is removed
    # from each argument.
    #
    # This function properly handles nested parenthesis and commas---these do not
    # define new arguments.
    # ----------------------------------------------------------------------

    def collect_args(self,tokenlist):
        args = []
        positions = []
        current_arg = []
        nesting = 1
        tokenlen = len(tokenlist)

        # Search for the opening '('.
        i = 0
        while (i < tokenlen) and (tokenlist[i].type in self.t_WS):
            i += 1

        if (i < tokenlen) and (tokenlist[i].value == '('):
            positions.append(i+1)
        else:
            self.error(self.source,tokenlist[0].lineno,"Missing '(' in macro arguments")
            return 0, [], []

        i += 1

        while i < tokenlen:
            t = tokenlist[i]
            if t.value == '(':
                current_arg.append(t)
                nesting += 1
            elif t.value == ')':
                nesting -= 1
                if nesting == 0:
                    if current_arg:
                        args.append(self.tokenstrip(current_arg))
                        positions.append(i)
                    return i+1,args,positions
                current_arg.append(t)
            elif t.value == ',' and nesting == 1:
                args.append(self.tokenstrip(current_arg))
                positions.append(i+1)
                current_arg = []
            else:
                current_arg.append(t)
            i += 1

        # Missing end argument
        self.error(self.source,tokenlist[-1].lineno,"Missing ')' in macro arguments")
        return 0, [],[]

    # ----------------------------------------------------------------------
    # macro_prescan()
    #
    # Examine the macro value (token sequence) and identify patch points
    # This is used to speed up macro expansion later on---we'll know
    # right away where to apply patches to the value to form the expansion
    # ----------------------------------------------------------------------

    def macro_prescan(self,macro):
        macro.patch     = []             # Standard macro arguments
        macro.str_patch = []             # String conversion expansion
        macro.var_comma_patch = []       # Variadic macro comma patch
        i = 0
        while i < len(macro.value):
            if macro.value[i].type == self.t_ID and macro.value[i].value in macro.arglist:
                argnum = macro.arglist.index(macro.value[i].value)
                # Conversion of argument to a string
                if i > 0 and macro.value[i-1].value == '#':
                    macro.value[i] = copy.copy(macro.value[i])
                    macro.value[i].type = self.t_STRING
                    del macro.value[i-1]
                    macro.str_patch.append((argnum,i-1))
                    continue
                # Concatenation
                elif (i > 0 and macro.value[i-1].value == '##'):
                    macro.patch.append(('c',argnum,i-1))
                    del macro.value[i-1]
                    continue
                elif ((i+1) < len(macro.value) and macro.value[i+1].value == '##'):
                    macro.patch.append(('c',argnum,i))
                    i += 1
                    continue
                # Standard expansion
                else:
                    macro.patch.append(('e',argnum,i))
            elif macro.value[i].value == '##':
                if macro.variadic and (i > 0) and (macro.value[i-1].value == ',') and \
                        ((i+1) < len(macro.value)) and (macro.value[i+1].type == self.t_ID) and \
                        (macro.value[i+1].value == macro.vararg):
                    macro.var_comma_patch.append(i-1)
            i += 1
        macro.patch.sort(key=lambda x: x[2],reverse=True)

    # ----------------------------------------------------------------------
    # macro_expand_args()
    #
    # Given a Macro and list of arguments (each a token list), this method
    # returns an expanded version of a macro.  The return value is a token sequence
    # representing the replacement macro tokens
    # ----------------------------------------------------------------------

    def macro_expand_args(self,macro,args):
        # Make a copy of the macro token sequence
        rep = [copy.copy(_x) for _x in macro.value]

        # Make string expansion patches.  These do not alter the length of the replacement sequence

        str_expansion = {}
        for argnum, i in macro.str_patch:
            if argnum not in str_expansion:
                str_expansion[argnum] = ('"%s"' % "".join([x.value for x in args[argnum]])).replace("\\","\\\\")
            rep[i] = copy.copy(rep[i])
            rep[i].value = str_expansion[argnum]

        # Make the variadic macro comma patch.  If the variadic macro argument is empty, we get rid
        comma_patch = False
        if macro.variadic and not args[-1]:
            for i in macro.var_comma_patch:
                rep[i] = None
                comma_patch = True

        # Make all other patches.   The order of these matters.  It is assumed that the patch list
        # has been sorted in reverse order of patch location since replacements will cause the
        # size of the replacement sequence to expand from the patch point.

        expanded = { }
        for ptype, argnum, i in macro.patch:
            # Concatenation.   Argument is left unexpanded
            if ptype == 'c':
                rep[i:i+1] = args[argnum]
            # Normal expansion.  Argument is macro expanded first
            elif ptype == 'e':
                if argnum not in expanded:
                    expanded[argnum] = self.expand_macros(args[argnum])
                rep[i:i+1] = expanded[argnum]

        # Get rid of removed comma if necessary
        if comma_patch:
            rep = [_i for _i in rep if _i]

        return rep


    # ----------------------------------------------------------------------
    # expand_macros()
    #
    # Given a list of tokens, this function performs macro expansion.
    # The expanded argument is a dictionary that contains macros already
    # expanded.  This is used to prevent infinite recursion.
    # ----------------------------------------------------------------------

    def expand_macros(self,tokens,expanded=None):
        if expanded is None:
            expanded = {}
        i = 0
        while i < len(tokens):
            t = tokens[i]
            if t.type == self.t_ID:
                if t.value in self.macros and t.value not in expanded:
                    # Yes, we found a macro match
                    expanded[t.value] = True

                    m = self.macros[t.value]
                    if not m.arglist:
                        # A simple macro
                        ex = self.expand_macros([copy.copy(_x) for _x in m.value],expanded)
                        for e in ex:
                            e.lineno = t.lineno
                        tokens[i:i+1] = ex
                        i += len(ex)
                    else:
                        # A macro with arguments
                        j = i + 1
                        while j < len(tokens) and tokens[j].type in self.t_WS:
                            j += 1
                        if tokens[j].value == '(':
                            tokcount,args,positions = self.collect_args(tokens[j:])
                            if not m.variadic and len(args) !=  len(m.arglist):
                                self.error(self.source,t.lineno,"Macro %s requires %d arguments" % (t.value,len(m.arglist)))
                                i = j + tokcount
                            elif m.variadic and len(args) < len(m.arglist)-1:
                                if len(m.arglist) > 2:
                                    self.error(self.source,t.lineno,"Macro %s must have at least %d arguments" % (t.value, len(m.arglist)-1))
                                else:
                                    self.error(self.source,t.lineno,"Macro %s must have at least %d argument" % (t.value, len(m.arglist)-1))
                                i = j + tokcount
                            else:
                                if m.variadic:
                                    if len(args) == len(m.arglist)-1:
                                        args.append([])
                                    else:
                                        args[len(m.arglist)-1] = tokens[j+positions[len(m.arglist)-1]:j+tokcount-1]
                                        del args[len(m.arglist):]

                                # Get macro replacement text
                                rep = self.macro_expand_args(m,args)
                                rep = self.expand_macros(rep,expanded)
                                for r in rep:
                                    r.lineno = t.lineno
                                tokens[i:j+tokcount] = rep
                                i += len(rep)
                    del expanded[t.value]
                    continue
                elif t.value == '__LINE__':
                    t.type = self.t_INTEGER
                    t.value = self.t_INTEGER_TYPE(t.lineno)

            i += 1
        return tokens

    # ----------------------------------------------------------------------
    # evalexpr()
    #
    # Evaluate an expression token sequence for the purposes of evaluating
    # integral expressions.
    # ----------------------------------------------------------------------

    def evalexpr(self,tokens):
        # tokens = tokenize(line)
        # Search for defined macros
        i = 0
        while i < len(tokens):
            if tokens[i].type == self.t_ID and tokens[i].value == 'defined':
                j = i + 1
                needparen = False
                result = "0L"
                while j < len(tokens):
                    if tokens[j].type in self.t_WS:
                        j += 1
                        continue
                    elif tokens[j].type == self.t_ID:
                        if tokens[j].value in self.macros:
                            result = "1L"
                        else:
                            result = "0L"
                        if not needparen: break
                    elif tokens[j].value == '(':
                        needparen = True
                    elif tokens[j].value == ')':
                        break
                    else:
                        self.error(self.source,tokens[i].lineno,"Malformed defined()")
                    j += 1
                tokens[i].type = self.t_INTEGER
                tokens[i].value = self.t_INTEGER_TYPE(result)
                del tokens[i+1:j+1]
            i += 1
        tokens = self.expand_macros(tokens)
        for i,t in enumerate(tokens):
            if t.type == self.t_ID:
                tokens[i] = copy.copy(t)
                tokens[i].type = self.t_INTEGER
                tokens[i].value = self.t_INTEGER_TYPE("0L")
            elif t.type == self.t_INTEGER:
                tokens[i] = copy.copy(t)
                # Strip off any trailing suffixes
                tokens[i].value = str(tokens[i].value)
                while tokens[i].value[-1] not in "0123456789abcdefABCDEF":
                    tokens[i].value = tokens[i].value[:-1]

        expr = "".join([str(x.value) for x in tokens])
        expr = expr.replace("&&"," and ")
        expr = expr.replace("||"," or ")
        expr = expr.replace("!"," not ")
        try:
            result = eval(expr)
        except Exception:
            self.error(self.source,tokens[0].lineno,"Couldn't evaluate expression")
            result = 0
        return result

    # ----------------------------------------------------------------------
    # parsegen()
    #
    # Parse an input string/
    # ----------------------------------------------------------------------
    def parsegen(self,input,source=None):

        # Replace trigraph sequences
        t = trigraph(input)
        lines = self.group_lines(t)

        if not source:
            source = ""

        self.define("__FILE__ \"%s\"" % source)

        self.source = source
        chunk = []
        enable = True
        iftrigger = False
        ifstack = []

        for x in lines:
            for i,tok in enumerate(x):
                if tok.type not in self.t_WS: break
            if tok.value == '#':
                # Preprocessor directive

                # insert necessary whitespace instead of eaten tokens
                for tok in x:
                    if tok.type in self.t_WS and '\n' in tok.value:
                        chunk.append(tok)

                dirtokens = self.tokenstrip(x[i+1:])
                if dirtokens:
                    name = dirtokens[0].value
                    args = self.tokenstrip(dirtokens[1:])
                else:
                    name = ""
                    args = []

                if name == 'define':
                    if enable:
                        for tok in self.expand_macros(chunk):
                            yield tok
                        chunk = []
                        self.define(args)
                elif name == 'include':
                    if enable:
                        for tok in self.expand_macros(chunk):
                            yield tok
                        chunk = []
                        oldfile = self.macros['__FILE__']
                        for tok in self.include(args):
                            yield tok
                        self.macros['__FILE__'] = oldfile
                        self.source = source
                elif name == 'undef':
                    if enable:
                        for tok in self.expand_macros(chunk):
                            yield tok
                        chunk = []
                        self.undef(args)
                elif name == 'ifdef':
                    ifstack.append((enable,iftrigger))
                    if enable:
                        if not args[0].value in self.macros:
                            enable = False
                            iftrigger = False
                        else:
                            iftrigger = True
                elif name == 'ifndef':
                    ifstack.append((enable,iftrigger))
                    if enable:
                        if args[0].value in self.macros:
                            enable = False
                            iftrigger = False
                        else:
                            iftrigger = True
                elif name == 'if':
                    ifstack.append((enable,iftrigger))
                    if enable:
                        result = self.evalexpr(args)
                        if not result:
                            enable = False
                            iftrigger = False
                        else:
                            iftrigger = True
                elif name == 'elif':
                    if ifstack:
                        if ifstack[-1][0]:     # We only pay attention if outer "if" allows this
                            if enable:         # If already true, we flip enable False
                                enable = False
                            elif not iftrigger:   # If False, but not triggered yet, we'll check expression
                                result = self.evalexpr(args)
                                if result:
                                    enable  = True
                                    iftrigger = True
                    else:
                        self.error(self.source,dirtokens[0].lineno,"Misplaced #elif")

                elif name == 'else':
                    if ifstack:
                        if ifstack[-1][0]:
                            if enable:
                                enable = False
                            elif not iftrigger:
                                enable = True
                                iftrigger = True
                    else:
                        self.error(self.source,dirtokens[0].lineno,"Misplaced #else")

                elif name == 'endif':
                    if ifstack:
                        enable,iftrigger = ifstack.pop()
                    else:
                        self.error(self.source,dirtokens[0].lineno,"Misplaced #endif")
                else:
                    # Unknown preprocessor directive
                    pass

            else:
                # Normal text
                if enable:
                    chunk.extend(x)

        for tok in self.expand_macros(chunk):
            yield tok
        chunk = []

    # ----------------------------------------------------------------------
    # include()
    #
    # Implementation of file-inclusion
    # ----------------------------------------------------------------------

    def include(self,tokens):
        # Try to extract the filename and then process an include file
        if not tokens:
            return
        if tokens:
            if tokens[0].value != '<' and tokens[0].type != self.t_STRING:
                tokens = self.expand_macros(tokens)

            if tokens[0].value == '<':
                # Include <...>
                i = 1
                while i < len(tokens):
                    if tokens[i].value == '>':
                        break
                    i += 1
                else:
                    print("Malformed #include <...>")
                    return
                filename = "".join([x.value for x in tokens[1:i]])
                path = self.path + [""] + self.temp_path
            elif tokens[0].type == self.t_STRING:
                filename = tokens[0].value[1:-1]
                path = self.temp_path + [""] + self.path
            else:
                print("Malformed #include statement")
                return
        for p in path:
            iname = os.path.join(p,filename)
            try:
                data = open(iname,"r").read()
                dname = os.path.dirname(iname)
                if dname:
                    self.temp_path.insert(0,dname)
                for tok in self.parsegen(data,filename):
                    yield tok
                if dname:
                    del self.temp_path[0]
                break
            except IOError:
                pass
        else:
            print("Couldn't find '%s'" % filename)

    # ----------------------------------------------------------------------
    # define()
    #
    # Define a new macro
    # ----------------------------------------------------------------------

    def define(self,tokens):
        if isinstance(tokens,STRING_TYPES):
            tokens = self.tokenize(tokens)

        linetok = tokens
        try:
            name = linetok[0]
            if len(linetok) > 1:
                mtype = linetok[1]
            else:
                mtype = None
            if not mtype:
                m = Macro(name.value,[])
                self.macros[name.value] = m
            elif mtype.type in self.t_WS:
                # A normal macro
                m = Macro(name.value,self.tokenstrip(linetok[2:]))
                self.macros[name.value] = m
            elif mtype.value == '(':
                # A macro with arguments
                tokcount, args, positions = self.collect_args(linetok[1:])
                variadic = False
                for a in args:
                    if variadic:
                        print("No more arguments may follow a variadic argument")
                        break
                    astr = "".join([str(_i.value) for _i in a])
                    if astr == "...":
                        variadic = True
                        a[0].type = self.t_ID
                        a[0].value = '__VA_ARGS__'
                        variadic = True
                        del a[1:]
                        continue
                    elif astr[-3:] == "..." and a[0].type == self.t_ID:
                        variadic = True
                        del a[1:]
                        # If, for some reason, "." is part of the identifier, strip off the name for the purposes
                        # of macro expansion
                        if a[0].value[-3:] == '...':
                            a[0].value = a[0].value[:-3]
                        continue
                    if len(a) > 1 or a[0].type != self.t_ID:
                        print("Invalid macro argument")
                        break
                else:
                    mvalue = self.tokenstrip(linetok[1+tokcount:])
                    i = 0
                    while i < len(mvalue):
                        if i+1 < len(mvalue):
                            if mvalue[i].type in self.t_WS and mvalue[i+1].value == '##':
                                del mvalue[i]
                                continue
                            elif mvalue[i].value == '##' and mvalue[i+1].type in self.t_WS:
                                del mvalue[i+1]
                        i += 1
                    m = Macro(name.value,mvalue,[x[0].value for x in args],variadic)
                    self.macro_prescan(m)
                    self.macros[name.value] = m
            else:
                print("Bad macro definition")
        except LookupError:
            print("Bad macro definition")

    # ----------------------------------------------------------------------
    # undef()
    #
    # Undefine a macro
    # ----------------------------------------------------------------------

    def undef(self,tokens):
        id = tokens[0].value
        try:
            del self.macros[id]
        except LookupError:
            pass

    # ----------------------------------------------------------------------
    # parse()
    #
    # Parse input text.
    # ----------------------------------------------------------------------
    def parse(self,input,source=None,ignore={}):
        self.ignore = ignore
        self.parser = self.parsegen(input,source)

    # ----------------------------------------------------------------------
    # token()
    #
    # Method to return individual tokens
    # ----------------------------------------------------------------------
    def token(self):
        try:
            while True:
                tok = next(self.parser)
                if tok.type not in self.ignore: return tok
        except StopIteration:
            self.parser = None
            return None

if __name__ == '__main__':
    import ply.lex as lex
    lexer = lex.lex()

    # Run a preprocessor
    import sys
    f = open(sys.argv[1])
    input = f.read()

    p = Preprocessor(lexer)
    p.parse(input,sys.argv[1])
    while True:
        tok = p.token()
        if not tok: break
        print(p.source, tok)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\test_dtls.py ===
from __future__ import annotations

import random
from contextlib import asynccontextmanager
from itertools import count
from typing import TYPE_CHECKING, NoReturn

import attrs
import pytest

from trio._tests.pytest_plugin import skip_if_optional_else_raise

try:
    import trustme
    from OpenSSL import SSL
except ImportError as error:
    skip_if_optional_else_raise(error)


import trio
import trio.testing
from trio import DTLSChannel, DTLSEndpoint
from trio.testing._fake_net import FakeNet, UDPPacket

from .._core._tests.tutil import binds_ipv6, gc_collect_harder, slow

if TYPE_CHECKING:
    from collections.abc import AsyncGenerator

ca = trustme.CA()
server_cert = ca.issue_cert("example.com")

server_ctx = SSL.Context(SSL.DTLS_METHOD)
server_cert.configure_cert(server_ctx)

client_ctx = SSL.Context(SSL.DTLS_METHOD)
ca.configure_trust(client_ctx)


parametrize_ipv6 = pytest.mark.parametrize(
    "ipv6",
    [False, pytest.param(True, marks=binds_ipv6)],
    ids=["ipv4", "ipv6"],
)


def endpoint(**kwargs: int | bool) -> DTLSEndpoint:
    ipv6 = kwargs.pop("ipv6", False)
    family = trio.socket.AF_INET6 if ipv6 else trio.socket.AF_INET
    sock = trio.socket.socket(type=trio.socket.SOCK_DGRAM, family=family)
    return DTLSEndpoint(sock, **kwargs)


@asynccontextmanager
async def dtls_echo_server(
    *,
    autocancel: bool = True,
    mtu: int | None = None,
    ipv6: bool = False,
) -> AsyncGenerator[tuple[DTLSEndpoint, tuple[str, int]], None]:
    with endpoint(ipv6=ipv6) as server:
        localhost = "::1" if ipv6 else "127.0.0.1"
        await server.socket.bind((localhost, 0))
        async with trio.open_nursery() as nursery:

            async def echo_handler(dtls_channel: DTLSChannel) -> None:
                print(
                    "echo handler started: "
                    f"server {dtls_channel.endpoint.socket.getsockname()!r} "
                    f"client {dtls_channel.peer_address!r}",
                )
                if mtu is not None:
                    dtls_channel.set_ciphertext_mtu(mtu)
                try:
                    print("server starting do_handshake")
                    await dtls_channel.do_handshake()
                    print("server finished do_handshake")
                    # no branch for leaving this for loop because we only leave
                    # a channel by cancellation.
                    async for packet in dtls_channel:  # pragma: no branch
                        print(f"echoing {packet!r} -> {dtls_channel.peer_address!r}")
                        await dtls_channel.send(packet)
                except trio.BrokenResourceError:  # pragma: no cover
                    print("echo handler channel broken")

            await nursery.start(server.serve, server_ctx, echo_handler)

            yield server, server.socket.getsockname()

            if autocancel:
                nursery.cancel_scope.cancel()


@parametrize_ipv6
async def test_smoke(ipv6: bool) -> None:
    async with dtls_echo_server(ipv6=ipv6) as (_server_endpoint, address):
        with endpoint(ipv6=ipv6) as client_endpoint:
            client_channel = client_endpoint.connect(address, client_ctx)
            with pytest.raises(trio.NeedHandshakeError):
                client_channel.get_cleartext_mtu()

            await client_channel.do_handshake()
            await client_channel.send(b"hello")
            assert await client_channel.receive() == b"hello"
            await client_channel.send(b"goodbye")
            assert await client_channel.receive() == b"goodbye"

            with pytest.raises(
                ValueError,
                match=r"^openssl doesn't support sending empty DTLS packets$",
            ):
                await client_channel.send(b"")

            client_channel.set_ciphertext_mtu(1234)
            cleartext_mtu_1234 = client_channel.get_cleartext_mtu()
            client_channel.set_ciphertext_mtu(4321)
            assert client_channel.get_cleartext_mtu() > cleartext_mtu_1234
            client_channel.set_ciphertext_mtu(1234)
            assert client_channel.get_cleartext_mtu() == cleartext_mtu_1234


@slow
async def test_handshake_over_terrible_network(
    autojump_clock: trio.testing.MockClock,
) -> None:
    HANDSHAKES = 100
    r = random.Random(0)
    fn = FakeNet()
    fn.enable()
    # avoid spurious timeouts on slow machines
    autojump_clock.autojump_threshold = 0.001

    async with dtls_echo_server() as (_, address):
        async with trio.open_nursery() as nursery:

            async def route_packet(packet: UDPPacket) -> None:
                while True:
                    op = r.choices(
                        ["deliver", "drop", "dupe", "delay"],
                        weights=[0.7, 0.1, 0.1, 0.1],
                    )[0]
                    print(f"{packet.source} -> {packet.destination}: {op}")
                    if op == "drop":
                        return
                    elif op == "dupe":
                        fn.send_packet(packet)
                    elif op == "delay":
                        await trio.sleep(r.random() * 3)
                    # I wanted to test random packet corruption too, but it turns out
                    # openssl has a bug in the following scenario:
                    #
                    # - client sends ClientHello
                    # - server sends HelloVerifyRequest with cookie -- but cookie is
                    #   invalid b/c either the ClientHello or HelloVerifyRequest was
                    #   corrupted
                    # - client re-sends ClientHello with invalid cookie
                    # - server replies with new HelloVerifyRequest and correct cookie
                    #
                    # At this point, the client *should* switch to the new, valid
                    # cookie. But OpenSSL doesn't; it stubbornly insists on re-sending
                    # the original, invalid cookie over and over. In theory we could
                    # work around this by detecting cookie changes and starting over
                    # with a whole new SSL object, but (a) it doesn't seem worth it, (b)
                    # when I tried then I ran into another issue where OpenSSL got stuck
                    # in an infinite loop sending alerts over and over, which I didn't
                    # dig into because see (a).
                    #
                    # elif op == "distort":
                    #     payload = bytearray(packet.payload)
                    #     payload[r.randrange(len(payload))] ^= 1 << r.randrange(8)
                    #     packet = attrs.evolve(packet, payload=payload)
                    else:
                        assert op == "deliver"
                        print(
                            f"{packet.source} -> {packet.destination}: delivered"
                            f" {packet.payload.hex()}",
                        )
                        fn.deliver_packet(packet)
                        break

            def route_packet_wrapper(packet: UDPPacket) -> None:
                try:  # noqa: SIM105  # suppressible-exception
                    nursery.start_soon(route_packet, packet)
                except RuntimeError:  # pragma: no cover
                    # We're exiting the nursery, so any remaining packets can just get
                    # dropped
                    pass

            fn.route_packet = route_packet_wrapper  # type: ignore[assignment]  # TODO: Fix FakeNet typing

            for i in range(HANDSHAKES):
                print("#" * 80)
                print("#" * 80)
                print("#" * 80)
                with endpoint() as client_endpoint:
                    client = client_endpoint.connect(address, client_ctx)
                    print("client starting do_handshake")
                    await client.do_handshake()
                    print("client finished do_handshake")
                    msg = str(i).encode()
                    # Make multiple attempts to send data, because the network might
                    # drop it
                    while True:
                        with trio.move_on_after(10) as cscope:
                            await client.send(msg)
                            assert await client.receive() == msg
                        if not cscope.cancelled_caught:
                            break


async def test_implicit_handshake() -> None:
    async with dtls_echo_server() as (_, address):
        with endpoint() as client_endpoint:
            client = client_endpoint.connect(address, client_ctx)

            # Implicit handshake
            await client.send(b"xyz")
            assert await client.receive() == b"xyz"


async def test_full_duplex() -> None:
    # Tests simultaneous send/receive, and also multiple methods implicitly invoking
    # do_handshake simultaneously.
    with endpoint() as server_endpoint, endpoint() as client_endpoint:
        await server_endpoint.socket.bind(("127.0.0.1", 0))
        async with trio.open_nursery() as server_nursery:

            async def handler(channel: DTLSChannel) -> None:
                async with trio.open_nursery() as nursery:
                    nursery.start_soon(channel.send, b"from server")
                    nursery.start_soon(channel.receive)

            await server_nursery.start(server_endpoint.serve, server_ctx, handler)

            client = client_endpoint.connect(
                server_endpoint.socket.getsockname(),
                client_ctx,
            )
            async with trio.open_nursery() as nursery:
                nursery.start_soon(client.send, b"from client")
                nursery.start_soon(client.receive)

            server_nursery.cancel_scope.cancel()


async def test_channel_closing() -> None:
    async with dtls_echo_server() as (_, address):
        with endpoint() as client_endpoint:
            client = client_endpoint.connect(address, client_ctx)
            await client.do_handshake()
            client.close()

            with pytest.raises(trio.ClosedResourceError):
                await client.send(b"abc")
            with pytest.raises(trio.ClosedResourceError):
                await client.receive()

            # close is idempotent
            client.close()
            # can also aclose
            await client.aclose()


async def test_serve_exits_cleanly_on_close() -> None:
    async with dtls_echo_server(autocancel=False) as (server_endpoint, _address):
        server_endpoint.close()
        # Testing that the nursery exits even without being cancelled
    # close is idempotent
    server_endpoint.close()


async def test_client_multiplex() -> None:
    async with dtls_echo_server() as (_, address1), dtls_echo_server() as (_, address2):
        with endpoint() as client_endpoint:
            client1 = client_endpoint.connect(address1, client_ctx)
            client2 = client_endpoint.connect(address2, client_ctx)

            await client1.send(b"abc")
            await client2.send(b"xyz")
            assert await client2.receive() == b"xyz"
            assert await client1.receive() == b"abc"

            client_endpoint.close()

            with pytest.raises(trio.ClosedResourceError):
                await client1.send(b"xxx")
            with pytest.raises(trio.ClosedResourceError):
                await client2.receive()
            with pytest.raises(trio.ClosedResourceError):
                client_endpoint.connect(address1, client_ctx)

            async def null_handler(_: object) -> None:  # pragma: no cover
                pass

            async with trio.open_nursery() as nursery:
                with pytest.raises(trio.ClosedResourceError):
                    await nursery.start(client_endpoint.serve, server_ctx, null_handler)


async def test_dtls_over_dgram_only() -> None:
    with trio.socket.socket() as s:
        with pytest.raises(ValueError, match=r"^DTLS requires a SOCK_DGRAM socket$"):
            DTLSEndpoint(s)


async def test_double_serve() -> None:
    async def null_handler(_: object) -> None:  # pragma: no cover
        pass

    with endpoint() as server_endpoint:
        await server_endpoint.socket.bind(("127.0.0.1", 0))
        async with trio.open_nursery() as nursery:
            await nursery.start(server_endpoint.serve, server_ctx, null_handler)
            with pytest.raises(trio.BusyResourceError):
                await nursery.start(server_endpoint.serve, server_ctx, null_handler)

            nursery.cancel_scope.cancel()

        async with trio.open_nursery() as nursery:
            await nursery.start(server_endpoint.serve, server_ctx, null_handler)
            nursery.cancel_scope.cancel()


async def test_connect_to_non_server(autojump_clock: trio.abc.Clock) -> None:
    fn = FakeNet()
    fn.enable()
    with endpoint() as client1, endpoint() as client2:
        await client1.socket.bind(("127.0.0.1", 0))
        # This should just time out
        with trio.move_on_after(100) as cscope:
            channel = client2.connect(client1.socket.getsockname(), client_ctx)
            await channel.do_handshake()
        assert cscope.cancelled_caught


async def test_incoming_buffer_overflow(autojump_clock: trio.abc.Clock) -> None:
    fn = FakeNet()
    fn.enable()
    for buffer_size in [10, 20]:
        async with dtls_echo_server() as (_, address):
            with endpoint(incoming_packets_buffer=buffer_size) as client_endpoint:
                assert client_endpoint.incoming_packets_buffer == buffer_size
                client = client_endpoint.connect(address, client_ctx)
                for i in range(buffer_size + 15):
                    await client.send(str(i).encode())
                    await trio.sleep(1)
                stats = client.statistics()
                assert stats.incoming_packets_dropped_in_trio == 15
                for i in range(buffer_size):
                    assert await client.receive() == str(i).encode()
                await client.send(b"buffer clear now")
                assert await client.receive() == b"buffer clear now"


async def test_server_socket_doesnt_crash_on_garbage(
    autojump_clock: trio.abc.Clock,
) -> None:
    fn = FakeNet()
    fn.enable()

    from trio._dtls import (
        ContentType,
        HandshakeFragment,
        HandshakeType,
        ProtocolVersion,
        Record,
        encode_handshake_fragment,
        encode_record,
    )

    client_hello = encode_record(
        Record(
            content_type=ContentType.handshake,
            version=ProtocolVersion.DTLS10,
            epoch_seqno=0,
            payload=encode_handshake_fragment(
                HandshakeFragment(
                    msg_type=HandshakeType.client_hello,
                    msg_len=10,
                    msg_seq=0,
                    frag_offset=0,
                    frag_len=10,
                    frag=bytes(10),
                ),
            ),
        ),
    )

    client_hello_extended = client_hello + b"\x00"
    client_hello_short = client_hello[:-1]
    # cuts off in middle of handshake message header
    client_hello_really_short = client_hello[:14]
    client_hello_corrupt_record_len = bytearray(client_hello)
    client_hello_corrupt_record_len[11] = 0xFF

    client_hello_fragmented = encode_record(
        Record(
            content_type=ContentType.handshake,
            version=ProtocolVersion.DTLS10,
            epoch_seqno=0,
            payload=encode_handshake_fragment(
                HandshakeFragment(
                    msg_type=HandshakeType.client_hello,
                    msg_len=20,
                    msg_seq=0,
                    frag_offset=0,
                    frag_len=10,
                    frag=bytes(10),
                ),
            ),
        ),
    )

    client_hello_trailing_data_in_record = encode_record(
        Record(
            content_type=ContentType.handshake,
            version=ProtocolVersion.DTLS10,
            epoch_seqno=0,
            payload=encode_handshake_fragment(
                HandshakeFragment(
                    msg_type=HandshakeType.client_hello,
                    msg_len=20,
                    msg_seq=0,
                    frag_offset=0,
                    frag_len=10,
                    frag=bytes(10),
                ),
            )
            + b"\x00",
        ),
    )

    handshake_empty = encode_record(
        Record(
            content_type=ContentType.handshake,
            version=ProtocolVersion.DTLS10,
            epoch_seqno=0,
            payload=b"",
        ),
    )

    client_hello_truncated_in_cookie = encode_record(
        Record(
            content_type=ContentType.handshake,
            version=ProtocolVersion.DTLS10,
            epoch_seqno=0,
            payload=bytes(2 + 32 + 1) + b"\xff",
        ),
    )

    async with dtls_echo_server() as (_, address):
        with trio.socket.socket(type=trio.socket.SOCK_DGRAM) as sock:
            for bad_packet in [
                b"",
                b"xyz",
                client_hello_extended,
                client_hello_short,
                client_hello_really_short,
                client_hello_corrupt_record_len,
                client_hello_fragmented,
                client_hello_trailing_data_in_record,
                handshake_empty,
                client_hello_truncated_in_cookie,
            ]:
                await sock.sendto(bad_packet, address)
                await trio.sleep(1)


async def test_invalid_cookie_rejected(autojump_clock: trio.abc.Clock) -> None:
    fn = FakeNet()
    fn.enable()

    from trio._dtls import BadPacket, decode_client_hello_untrusted

    with trio.CancelScope() as cscope:
        # the first 11 bytes of ClientHello aren't protected by the cookie, so only test
        # corrupting bytes after that.
        offset_to_corrupt = count(11)

        def route_packet(packet: UDPPacket) -> None:
            try:
                _, cookie, _ = decode_client_hello_untrusted(packet.payload)
            except BadPacket:
                pass
            else:
                if len(cookie) != 0:
                    # this is a challenge response packet
                    # let's corrupt the next offset so the handshake should fail
                    payload = bytearray(packet.payload)
                    offset = next(offset_to_corrupt)
                    if offset >= len(payload):
                        # We've tried all offsets. Clamp offset to the end of the
                        # payload, and terminate the test.
                        offset = len(payload) - 1
                        cscope.cancel()
                    payload[offset] ^= 0x01
                    packet = attrs.evolve(packet, payload=payload)

            fn.deliver_packet(packet)

        fn.route_packet = route_packet  # type: ignore[assignment]  # TODO: Fix FakeNet typing

        async with dtls_echo_server() as (_, address):
            while True:
                with endpoint() as client:
                    channel = client.connect(address, client_ctx)
                    await channel.do_handshake()
    assert cscope.cancelled_caught


async def test_client_cancels_handshake_and_starts_new_one(
    autojump_clock: trio.abc.Clock,
) -> None:
    # if a client disappears during the handshake, and then starts a new handshake from
    # scratch, then the first handler's channel should fail, and a new handler get
    # started
    fn = FakeNet()
    fn.enable()

    with endpoint() as server, endpoint() as client:
        await server.socket.bind(("127.0.0.1", 0))
        async with trio.open_nursery() as nursery:
            first_time = True

            async def handler(channel: DTLSChannel) -> None:
                nonlocal first_time
                if first_time:
                    first_time = False
                    print("handler: first time, cancelling connect")
                    connect_cscope.cancel()
                    await trio.sleep(0.5)
                    print("handler: handshake should fail now")
                    with pytest.raises(trio.BrokenResourceError):
                        await channel.do_handshake()
                else:
                    print("handler: not first time, sending hello")
                    await channel.send(b"hello")

            await nursery.start(server.serve, server_ctx, handler)

            print("client: starting first connect")
            with trio.CancelScope() as connect_cscope:
                channel = client.connect(server.socket.getsockname(), client_ctx)
                await channel.do_handshake()
            assert connect_cscope.cancelled_caught

            print("client: starting second connect")
            channel = client.connect(server.socket.getsockname(), client_ctx)
            assert await channel.receive() == b"hello"

            # Give handlers a chance to finish
            await trio.sleep(10)
            nursery.cancel_scope.cancel()


async def test_swap_client_server() -> None:
    with endpoint() as a, endpoint() as b:
        await a.socket.bind(("127.0.0.1", 0))
        await b.socket.bind(("127.0.0.1", 0))

        async def echo_handler(channel: DTLSChannel) -> None:
            async for packet in channel:
                await channel.send(packet)

        async def crashing_echo_handler(channel: DTLSChannel) -> None:
            with pytest.raises(trio.BrokenResourceError):
                await echo_handler(channel)

        async with trio.open_nursery() as nursery:
            await nursery.start(a.serve, server_ctx, crashing_echo_handler)
            await nursery.start(b.serve, server_ctx, echo_handler)

            b_to_a = b.connect(a.socket.getsockname(), client_ctx)
            await b_to_a.send(b"b as client")
            assert await b_to_a.receive() == b"b as client"

            a_to_b = a.connect(b.socket.getsockname(), client_ctx)
            await a_to_b.do_handshake()
            with pytest.raises(trio.BrokenResourceError):
                await b_to_a.send(b"association broken")
            await a_to_b.send(b"a as client")
            assert await a_to_b.receive() == b"a as client"

            nursery.cancel_scope.cancel()


@slow
async def test_openssl_retransmit_doesnt_break_stuff() -> None:
    # can't use autojump_clock here, because the point of the test is to wait for
    # openssl's built-in retransmit timer to expire, which is hard-coded to use
    # wall-clock time.
    fn = FakeNet()
    fn.enable()

    blackholed = True

    def route_packet(packet: UDPPacket) -> None:
        if blackholed:
            print("dropped packet", packet)
            return
        print("delivered packet", packet)
        # packets.append(
        #     scapy.all.IP(
        #         src=packet.source.ip.compressed, dst=packet.destination.ip.compressed
        #     )
        #     / scapy.all.UDP(sport=packet.source.port, dport=packet.destination.port)
        #     / packet.payload
        # )
        fn.deliver_packet(packet)

    fn.route_packet = route_packet  # type: ignore[assignment]  # TODO add type annotations for FakeNet

    async with dtls_echo_server() as (server_endpoint, address):
        with endpoint() as client_endpoint:
            async with trio.open_nursery() as nursery:

                async def connecter() -> None:
                    client = client_endpoint.connect(address, client_ctx)
                    await client.do_handshake(initial_retransmit_timeout=1.5)
                    await client.send(b"hi")
                    assert await client.receive() == b"hi"

                nursery.start_soon(connecter)

                # openssl's default timeout is 1 second, so this ensures that it thinks
                # the timeout has expired
                await trio.sleep(1.1)
                # disable blackholing and send a garbage packet to wake up openssl so it
                # notices the timeout has expired
                blackholed = False
                await server_endpoint.socket.sendto(
                    b"xxx",
                    client_endpoint.socket.getsockname(),
                )
                # now the client task should finish connecting and exit cleanly

    # scapy.all.wrpcap("/tmp/trace.pcap", packets)


async def test_initial_retransmit_timeout_configuration(
    autojump_clock: trio.abc.Clock,
) -> None:
    fn = FakeNet()
    fn.enable()

    blackholed = True

    def route_packet(packet: UDPPacket) -> None:
        nonlocal blackholed
        if blackholed:
            blackholed = False
        else:
            fn.deliver_packet(packet)

    fn.route_packet = route_packet  # type: ignore[assignment]  # TODO add type annotations for FakeNet

    async with dtls_echo_server() as (_, address):
        for t in [1, 2, 4]:
            with endpoint() as client:
                before = trio.current_time()
                blackholed = True
                channel = client.connect(address, client_ctx)
                await channel.do_handshake(initial_retransmit_timeout=t)
                after = trio.current_time()
                assert after - before == t


async def test_explicit_tiny_mtu_is_respected() -> None:
    # ClientHello is ~240 bytes, and it can't be fragmented, so our mtu has to
    # be larger than that. (300 is still smaller than any real network though.)
    MTU = 300

    fn = FakeNet()
    fn.enable()

    def route_packet(packet: UDPPacket) -> None:
        print(f"delivering {packet}")
        print(f"payload size: {len(packet.payload)}")
        assert len(packet.payload) <= MTU
        fn.deliver_packet(packet)

    fn.route_packet = route_packet  # type: ignore[assignment]  # TODO add type annotations for FakeNet

    async with dtls_echo_server(mtu=MTU) as (_server, address):
        with endpoint() as client:
            channel = client.connect(address, client_ctx)
            channel.set_ciphertext_mtu(MTU)
            await channel.do_handshake()
            await channel.send(b"hi")
            assert await channel.receive() == b"hi"


@parametrize_ipv6
async def test_handshake_handles_minimum_network_mtu(
    ipv6: bool,
    autojump_clock: trio.abc.Clock,
) -> None:
    # Fake network that has the minimum allowable MTU for whatever protocol we're using.
    fn = FakeNet()
    fn.enable()

    mtu = 1280 - 48 if ipv6 else 576 - 28

    def route_packet(packet: UDPPacket) -> None:
        if len(packet.payload) > mtu:
            print(f"dropping {packet}")
        else:
            print(f"delivering {packet}")
            fn.deliver_packet(packet)

    fn.route_packet = route_packet  # type: ignore[assignment]  # TODO add type annotations for FakeNet

    # See if we can successfully do a handshake -- some of the volleys will get dropped,
    # and the retransmit logic should detect this and back off the MTU to something
    # smaller until it succeeds.
    async with dtls_echo_server(ipv6=ipv6) as (_, address):
        with endpoint(ipv6=ipv6) as client_endpoint:
            client = client_endpoint.connect(address, client_ctx)
            # the handshake mtu backoff shouldn't affect the return value from
            # get_cleartext_mtu, b/c that's under the user's control via
            # set_ciphertext_mtu
            client.set_ciphertext_mtu(9999)
            await client.send(b"xyz")
            assert await client.receive() == b"xyz"
            assert client.get_cleartext_mtu() > 9000  # as vegeta said


@pytest.mark.filterwarnings("always:unclosed DTLS:ResourceWarning")
async def test_system_task_cleaned_up_on_gc() -> None:
    before_tasks = trio.lowlevel.current_statistics().tasks_living

    # We put this into a sub-function so that everything automatically becomes garbage
    # when the frame exits. For some reason just doing 'del e' wasn't enough on pypy
    # with coverage enabled -- I think we were hitting this bug:
    #     https://foss.heptapod.net/pypy/pypy/-/issues/3656
    async def start_and_forget_endpoint() -> int:
        e = endpoint()

        # This connection/handshake attempt can't succeed. The only purpose is to force
        # the endpoint to set up a receive loop.
        with trio.socket.socket(type=trio.socket.SOCK_DGRAM) as s:
            await s.bind(("127.0.0.1", 0))
            c = e.connect(s.getsockname(), client_ctx)
            async with trio.open_nursery() as nursery:
                nursery.start_soon(c.do_handshake)
                await trio.testing.wait_all_tasks_blocked()
                nursery.cancel_scope.cancel()

        during_tasks = trio.lowlevel.current_statistics().tasks_living
        return during_tasks

    with pytest.warns(ResourceWarning):  # noqa: PT031
        during_tasks = await start_and_forget_endpoint()
        await trio.testing.wait_all_tasks_blocked()
        gc_collect_harder()

    await trio.testing.wait_all_tasks_blocked()

    after_tasks = trio.lowlevel.current_statistics().tasks_living
    assert before_tasks < during_tasks
    assert before_tasks == after_tasks


@pytest.mark.filterwarnings("always:unclosed DTLS:ResourceWarning")
async def test_gc_before_system_task_starts() -> None:
    e = endpoint()

    with pytest.warns(ResourceWarning):  # noqa: PT031
        del e
        gc_collect_harder()

    await trio.testing.wait_all_tasks_blocked()


@pytest.mark.filterwarnings("always:unclosed DTLS:ResourceWarning")
async def test_gc_as_packet_received() -> None:
    fn = FakeNet()
    fn.enable()

    e = endpoint()
    await e.socket.bind(("127.0.0.1", 0))
    e._ensure_receive_loop()

    await trio.testing.wait_all_tasks_blocked()

    with trio.socket.socket(type=trio.socket.SOCK_DGRAM) as s:
        await s.sendto(b"xxx", e.socket.getsockname())
    # At this point, the endpoint's receive loop has been marked runnable because it
    # just received a packet; closing the endpoint socket won't interrupt that. But by
    # the time it wakes up to process the packet, the endpoint will be gone.
    with pytest.warns(ResourceWarning):  # noqa: PT031
        del e
        gc_collect_harder()


@pytest.mark.filterwarnings("always:unclosed DTLS:ResourceWarning")
def test_gc_after_trio_exits() -> None:
    async def main() -> DTLSEndpoint:
        # We use fakenet just to make sure no real sockets can leak out of the test
        # case - on pypy somehow the socket was outliving the gc_collect_harder call
        # below. Since the test is just making sure DTLSEndpoint.__del__ doesn't explode
        # when called after trio exits, it doesn't need a real socket.
        fn = FakeNet()
        fn.enable()
        return endpoint()

    e = trio.run(main)
    with pytest.warns(ResourceWarning):  # noqa: PT031
        del e
        gc_collect_harder()


async def test_already_closed_socket_doesnt_crash() -> None:
    with endpoint() as e:
        # We close the socket before checkpointing, so the socket will already be closed
        # when the system task starts up
        e.socket.close()
        # Now give it a chance to start up, and hopefully not crash
        await trio.testing.wait_all_tasks_blocked()


async def test_socket_closed_while_processing_clienthello(
    autojump_clock: trio.abc.Clock,
) -> None:
    fn = FakeNet()
    fn.enable()

    # Check what happens if the socket is discovered to be closed when sending a
    # HelloVerifyRequest, since that has its own sending logic
    async with dtls_echo_server() as (server, address):

        def route_packet(packet: UDPPacket) -> None:
            fn.deliver_packet(packet)
            server.socket.close()

        fn.route_packet = route_packet  # type: ignore[assignment]  # TODO add type annotations for FakeNet

        with endpoint() as client_endpoint:
            with trio.move_on_after(10):
                client = client_endpoint.connect(address, client_ctx)
                await client.do_handshake()


async def test_association_replaced_while_handshake_running(
    autojump_clock: trio.abc.Clock,
) -> None:
    fn = FakeNet()
    fn.enable()

    def route_packet(packet: UDPPacket) -> None:
        pass

    fn.route_packet = route_packet  # type: ignore[assignment]  # TODO add type annotations for FakeNet

    async with dtls_echo_server() as (_, address):
        with endpoint() as client_endpoint:
            c1 = client_endpoint.connect(address, client_ctx)
            async with trio.open_nursery() as nursery:

                async def doomed_handshake() -> None:
                    with pytest.raises(trio.BrokenResourceError):
                        await c1.do_handshake()

                nursery.start_soon(doomed_handshake)

                await trio.sleep(10)

                client_endpoint.connect(address, client_ctx)


async def test_association_replaced_before_handshake_starts() -> None:
    fn = FakeNet()
    fn.enable()

    # This test shouldn't send any packets
    def route_packet(packet: UDPPacket) -> NoReturn:  # pragma: no cover
        raise AssertionError()

    fn.route_packet = route_packet  # type: ignore[assignment]  # TODO add type annotations for FakeNet

    async with dtls_echo_server() as (_, address):
        with endpoint() as client_endpoint:
            c1 = client_endpoint.connect(address, client_ctx)
            client_endpoint.connect(address, client_ctx)
            with pytest.raises(trio.BrokenResourceError):
                await c1.do_handshake()


async def test_send_to_closed_local_port() -> None:
    # On Windows, sending a UDP packet to a closed local port can cause a weird
    # ECONNRESET error later, inside the receive task. Make sure we're handling it
    # properly.
    async with dtls_echo_server() as (_, address):
        with endpoint() as client_endpoint:
            async with trio.open_nursery() as nursery:
                for i in range(1, 10):
                    channel = client_endpoint.connect(("127.0.0.1", i), client_ctx)
                    nursery.start_soon(channel.do_handshake)
                channel = client_endpoint.connect(address, client_ctx)
                await channel.send(b"xxx")
                assert await channel.receive() == b"xxx"
                nursery.cancel_scope.cancel()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pyreadline3\console\console.py ===
# -*- coding: utf-8 -*-
# *****************************************************************************
#       Copyright (C) 2003-2006 Gary Bishop.
#       Copyright (C) 2006-2020 Jorgen Stenarson. <jorgen.stenarson@bostream.nu>
#       Copyright (C) 2020 Bassem Girgis. <brgirgis@gmail.com>
#
#  Distributed under the terms of the BSD License.  The full license is in
#  the file COPYING, distributed as part of this software.
# *****************************************************************************


from .event import Event

"""Cursor control and color for the Windows console.

This was modeled after the C extension of the same name by Fredrik Lundh.
"""

# primitive debug printing that won't interfere with the screen

import os
import re
import sys
import traceback

import pyreadline3.unicode_helper as unicode_helper
from pyreadline3.console.ansi import AnsiState, AnsiWriter
from pyreadline3.keysyms import KeyPress, make_KeyPress
from pyreadline3.logger import log
from pyreadline3.unicode_helper import ensure_str, ensure_unicode

try:
    import ctypes.util
    from ctypes import *
    from ctypes.wintypes import *

    from _ctypes import call_function
except ImportError:
    raise ImportError("You need ctypes to run this code")


def nolog(string):
    pass


log = nolog


# some constants we need
STD_INPUT_HANDLE = -10
STD_OUTPUT_HANDLE = -11
ENABLE_WINDOW_INPUT = 0x0008
ENABLE_MOUSE_INPUT = 0x0010
ENABLE_PROCESSED_INPUT = 0x0001
WHITE = 0x7
BLACK = 0
MENU_EVENT = 0x0008
KEY_EVENT = 0x0001
MOUSE_MOVED = 0x0001
MOUSE_EVENT = 0x0002
WINDOW_BUFFER_SIZE_EVENT = 0x0004
FOCUS_EVENT = 0x0010
MENU_EVENT = 0x0008
VK_SHIFT = 0x10
VK_CONTROL = 0x11
VK_MENU = 0x12
GENERIC_READ = int(0x80000000)
GENERIC_WRITE = 0x40000000

# Windows structures we'll need later


class COORD(Structure):
    _fields_ = [("X", c_short), ("Y", c_short)]


class SMALL_RECT(Structure):
    _fields_ = [
        ("Left", c_short),
        ("Top", c_short),
        ("Right", c_short),
        ("Bottom", c_short),
    ]


class CONSOLE_SCREEN_BUFFER_INFO(Structure):
    _fields_ = [
        ("dwSize", COORD),
        ("dwCursorPosition", COORD),
        ("wAttributes", c_short),
        ("srWindow", SMALL_RECT),
        ("dwMaximumWindowSize", COORD),
    ]


class CHAR_UNION(Union):
    _fields_ = [("UnicodeChar", c_wchar), ("AsciiChar", c_char)]


class CHAR_INFO(Structure):
    _fields_ = [("Char", CHAR_UNION), ("Attributes", c_short)]


class KEY_EVENT_RECORD(Structure):
    _fields_ = [
        ("bKeyDown", c_byte),
        ("pad2", c_byte),
        ("pad1", c_short),
        ("wRepeatCount", c_short),
        ("wVirtualKeyCode", c_short),
        ("wVirtualScanCode", c_short),
        ("uChar", CHAR_UNION),
        ("dwControlKeyState", c_int),
    ]


class MOUSE_EVENT_RECORD(Structure):
    _fields_ = [
        ("dwMousePosition", COORD),
        ("dwButtonState", c_int),
        ("dwControlKeyState", c_int),
        ("dwEventFlags", c_int),
    ]


class WINDOW_BUFFER_SIZE_RECORD(Structure):
    _fields_ = [("dwSize", COORD)]


class MENU_EVENT_RECORD(Structure):
    _fields_ = [("dwCommandId", c_uint)]


class FOCUS_EVENT_RECORD(Structure):
    _fields_ = [("bSetFocus", c_byte)]


class INPUT_UNION(Union):
    _fields_ = [
        ("KeyEvent", KEY_EVENT_RECORD),
        ("MouseEvent", MOUSE_EVENT_RECORD),
        ("WindowBufferSizeEvent", WINDOW_BUFFER_SIZE_RECORD),
        ("MenuEvent", MENU_EVENT_RECORD),
        ("FocusEvent", FOCUS_EVENT_RECORD),
    ]


class INPUT_RECORD(Structure):
    _fields_ = [("EventType", c_short), ("Event", INPUT_UNION)]


class CONSOLE_CURSOR_INFO(Structure):
    _fields_ = [("dwSize", c_int), ("bVisible", c_byte)]


# I didn't want to have to individually import these so I made a list, they are
# added to the Console class later in this file.

funcs = [
    "AllocConsole",
    "CreateConsoleScreenBuffer",
    "FillConsoleOutputAttribute",
    "FillConsoleOutputCharacterW",
    "FreeConsole",
    "GetConsoleCursorInfo",
    "GetConsoleMode",
    "GetConsoleScreenBufferInfo",
    "GetConsoleTitleW",
    "GetProcAddress",
    "GetStdHandle",
    "PeekConsoleInputW",
    "ReadConsoleInputW",
    "ScrollConsoleScreenBufferW",
    "SetConsoleActiveScreenBuffer",
    "SetConsoleCursorInfo",
    "SetConsoleCursorPosition",
    "SetConsoleMode",
    "SetConsoleScreenBufferSize",
    "SetConsoleTextAttribute",
    "SetConsoleTitleW",
    "SetConsoleWindowInfo",
    "WriteConsoleW",
    "WriteConsoleOutputCharacterW",
    "WriteFile",
]

# I don't want events for these keys, they are just a bother for my application
key_modifiers = {
    VK_SHIFT: 1,
    VK_CONTROL: 1,
    VK_MENU: 1,  # alt key
    0x5B: 1,  # windows key
}


def split_block(text, size=1000):
    return [text[start : start + size] for start in range(0, len(text), size)]


class Console(object):
    """Console driver for Windows."""

    def __init__(self, newbuffer=0):
        """Initialize the Console object.

        newbuffer=1 will allocate a new buffer so the old content will be restored
        on exit.
        """
        # Do I need the following line? It causes a console to be created whenever
        # readline is imported into a pythonw application which seems wrong. Things
        # seem to work without it...
        # self.AllocConsole()

        if newbuffer:
            self.hout = self.CreateConsoleScreenBuffer(
                GENERIC_READ | GENERIC_WRITE, 0, None, 1, None
            )
            self.SetConsoleActiveScreenBuffer(self.hout)
        else:
            self.hout = self.GetStdHandle(STD_OUTPUT_HANDLE)

        self.hin = self.GetStdHandle(STD_INPUT_HANDLE)
        self.inmode = DWORD(0)
        self.GetConsoleMode(self.hin, byref(self.inmode))
        self.SetConsoleMode(self.hin, 0xF)
        info = CONSOLE_SCREEN_BUFFER_INFO()
        self.GetConsoleScreenBufferInfo(self.hout, byref(info))
        self.attr = info.wAttributes
        self.saveattr = info.wAttributes  # remember the initial colors
        self.defaultstate = AnsiState()
        self.defaultstate.winattr = info.wAttributes
        self.ansiwriter = AnsiWriter(self.defaultstate)

        background = self.attr & 0xF0
        for escape in self.escape_to_color:
            if self.escape_to_color[escape] is not None:
                self.escape_to_color[escape] |= background
        log("initial attr=%x" % self.attr)
        self.softspace = 0  # this is for using it as a file-like object
        self.serial = 0

        self.pythondll = ctypes.pythonapi
        self.inputHookPtr = c_void_p.from_address(
            addressof(self.pythondll.PyOS_InputHook)
        ).value

        self.pythondll.PyMem_RawMalloc.restype = c_size_t
        self.pythondll.PyMem_RawMalloc.argtypes = [c_size_t]
        setattr(Console, "PyMem_Malloc", self.pythondll.PyMem_RawMalloc)

    def __del__(self):
        """Cleanup the console when finished."""
        # I don't think this ever gets called
        self.SetConsoleTextAttribute(self.hout, self.saveattr)
        self.SetConsoleMode(self.hin, self.inmode)
        self.FreeConsole()

    def _get_top_bot(self):
        info = CONSOLE_SCREEN_BUFFER_INFO()
        self.GetConsoleScreenBufferInfo(self.hout, byref(info))
        rect = info.srWindow
        top = rect.Top
        bot = rect.Bottom
        return top, bot

    def fixcoord(self, x, y):
        """Return a long with x and y packed inside,
        also handle negative x and y."""
        if x < 0 or y < 0:
            info = CONSOLE_SCREEN_BUFFER_INFO()
            self.GetConsoleScreenBufferInfo(self.hout, byref(info))
            if x < 0:
                x = info.srWindow.Right - x
                y = info.srWindow.Bottom + y

        # this is a hack! ctypes won't pass structures but COORD is
        # just like a long, so this works.
        return c_int(y << 16 | x)

    def pos(self, x=None, y=None):
        """Move or query the window cursor."""
        if x is None:
            info = CONSOLE_SCREEN_BUFFER_INFO()
            self.GetConsoleScreenBufferInfo(self.hout, byref(info))
            return (info.dwCursorPosition.X, info.dwCursorPosition.Y)
        else:
            return self.SetConsoleCursorPosition(self.hout, self.fixcoord(x, y))

    def home(self):
        """Move to home."""
        self.pos(0, 0)

    # Map ANSI color escape sequences into Windows Console Attributes

    terminal_escape = re.compile("(\001?\033\\[[0-9;]+m\002?)")
    escape_parts = re.compile("\001?\033\\[([0-9;]+)m\002?")
    escape_to_color = {
        "0;30": 0x0,  # black
        "0;31": 0x4,  # red
        "0;32": 0x2,  # green
        "0;33": 0x4 + 0x2,  # brown?
        "0;34": 0x1,  # blue
        "0;35": 0x1 + 0x4,  # purple
        "0;36": 0x2 + 0x4,  # cyan
        "0;37": 0x1 + 0x2 + 0x4,  # grey
        "1;30": 0x1 + 0x2 + 0x4,  # dark gray
        "1;31": 0x4 + 0x8,  # red
        "1;32": 0x2 + 0x8,  # light green
        "1;33": 0x4 + 0x2 + 0x8,  # yellow
        "1;34": 0x1 + 0x8,  # light blue
        "1;35": 0x1 + 0x4 + 0x8,  # light purple
        "1;36": 0x1 + 0x2 + 0x8,  # light cyan
        "1;37": 0x1 + 0x2 + 0x4 + 0x8,  # white
        "0": None,
    }

    # This pattern should match all characters that change the cursor position differently
    # than a normal character.
    motion_char_re = re.compile("([\n\r\t\010\007])")

    def write_scrolling(self, text, attr=None):
        """write text at current cursor position while watching for scrolling.

        If the window scrolls because you are at the bottom of the screen
        buffer, all positions that you are storing will be shifted by the
        scroll amount. For example, I remember the cursor position of the
        prompt so that I can redraw the line but if the window scrolls,
        the remembered position is off.

        This variant of write tries to keep track of the cursor position
        so that it will know when the screen buffer is scrolled. It
        returns the number of lines that the buffer scrolled.

        """
        text = ensure_unicode(text)
        x, y = self.pos()
        w, h = self.size()
        scroll = 0  # the result
        # split the string into ordinary characters and funny characters
        chunks = self.motion_char_re.split(text)
        for chunk in chunks:
            n = self.write_color(chunk, attr)
            if len(chunk) == 1:  # the funny characters will be alone
                if chunk[0] == "\n":  # newline
                    x = 0
                    y += 1
                elif chunk[0] == "\r":  # carriage return
                    x = 0
                elif chunk[0] == "\t":  # tab
                    x = 8 * (int(x / 8) + 1)
                    if x > w:  # newline
                        x -= w
                        y += 1
                elif chunk[0] == "\007":  # bell
                    pass
                elif chunk[0] == "\010":
                    x -= 1
                    if x < 0:
                        y -= 1  # backed up 1 line
                else:  # ordinary character
                    x += 1
                if x == w:  # wrap
                    x = 0
                    y += 1
                if y == h:  # scroll
                    scroll += 1
                    y = h - 1
            else:  # chunk of ordinary characters
                x += n
                l = int(x / w)  # lines we advanced
                x = x % w  # new x value
                y += l
                if y >= h:  # scroll
                    scroll += y - h + 1
                    y = h - 1
        return scroll

    def write_color(self, text, attr=None):
        text = ensure_unicode(text)
        n, res = self.ansiwriter.write_color(text, attr)
        junk = DWORD(0)
        for attr, chunk in res:
            log("console.attr:%s" % (attr))
            log("console.chunk:%s" % (chunk))
            self.SetConsoleTextAttribute(self.hout, attr.winattr)
            for short_chunk in split_block(chunk):
                self.WriteConsoleW(
                    self.hout, short_chunk, len(short_chunk), byref(junk), None
                )
        return n

    def write_plain(self, text, attr=None):
        """write text at current cursor position."""
        text = ensure_unicode(text)
        log('write("%s", %s)' % (text, attr))
        if attr is None:
            attr = self.attr
        junk = DWORD(0)
        self.SetConsoleTextAttribute(self.hout, attr)
        for short_chunk in split_block(chunk):
            self.WriteConsoleW(
                self.hout,
                ensure_unicode(short_chunk),
                len(short_chunk),
                byref(junk),
                None,
            )
        return len(text)

    # This function must be used to ensure functioning with EMACS
    # Emacs sets the EMACS environment variable
    if "EMACS" in os.environ:

        def write_color(self, text, attr=None):
            text = ensure_str(text)
            junk = DWORD(0)
            self.WriteFile(self.hout, text, len(text), byref(junk), None)
            return len(text)

        write_plain = write_color

    # make this class look like a file object
    def write(self, text):
        text = ensure_unicode(text)
        log('write("%s")' % text)
        return self.write_color(text)

    # write = write_scrolling

    def isatty(self):
        return True

    def flush(self):
        pass

    def page(self, attr=None, fill=" "):
        """Fill the entire screen."""
        if attr is None:
            attr = self.attr
        if len(fill) != 1:
            raise ValueError
        info = CONSOLE_SCREEN_BUFFER_INFO()
        self.GetConsoleScreenBufferInfo(self.hout, byref(info))
        if info.dwCursorPosition.X != 0 or info.dwCursorPosition.Y != 0:
            self.SetConsoleCursorPosition(self.hout, self.fixcoord(0, 0))

        w = info.dwSize.X
        n = DWORD(0)
        for y in range(info.dwSize.Y):
            self.FillConsoleOutputAttribute(
                self.hout, attr, w, self.fixcoord(0, y), byref(n)
            )
            self.FillConsoleOutputCharacterW(
                self.hout, ord(fill[0]), w, self.fixcoord(0, y), byref(n)
            )

        self.attr = attr

    def text(self, x, y, text, attr=None):
        """Write text at the given position."""
        if attr is None:
            attr = self.attr

        pos = self.fixcoord(x, y)
        n = DWORD(0)
        self.WriteConsoleOutputCharacterW(self.hout, text, len(text), pos, byref(n))
        self.FillConsoleOutputAttribute(self.hout, attr, n, pos, byref(n))

    def clear_to_end_of_window(self):
        top, bot = self._get_top_bot()
        pos = self.pos()
        w, h = self.size()
        self.rectangle((pos[0], pos[1], w, pos[1] + 1))
        if pos[1] < bot:
            self.rectangle((0, pos[1] + 1, w, bot + 1))

    def rectangle(self, rect, attr=None, fill=" "):
        """Fill Rectangle."""
        x0, y0, x1, y1 = rect
        n = DWORD(0)
        if attr is None:
            attr = self.attr
        for y in range(y0, y1):
            pos = self.fixcoord(x0, y)
            self.FillConsoleOutputAttribute(self.hout, attr, x1 - x0, pos, byref(n))
            self.FillConsoleOutputCharacterW(
                self.hout, ord(fill[0]), x1 - x0, pos, byref(n)
            )

    def scroll(self, rect, dx, dy, attr=None, fill=" "):
        """Scroll a rectangle."""
        if attr is None:
            attr = self.attr
        x0, y0, x1, y1 = rect
        source = SMALL_RECT(x0, y0, x1 - 1, y1 - 1)
        dest = self.fixcoord(x0 + dx, y0 + dy)
        style = CHAR_INFO()
        style.Char.AsciiChar = ensure_str(fill[0])
        style.Attributes = attr

        return self.ScrollConsoleScreenBufferW(
            self.hout, byref(source), byref(source), dest, byref(style)
        )

    def scroll_window(self, lines):
        """Scroll the window by the indicated number of lines."""
        info = CONSOLE_SCREEN_BUFFER_INFO()
        self.GetConsoleScreenBufferInfo(self.hout, byref(info))
        rect = info.srWindow
        log("sw: rtop=%d rbot=%d" % (rect.Top, rect.Bottom))
        top = rect.Top + lines
        bot = rect.Bottom + lines
        h = bot - top
        maxbot = info.dwSize.Y - 1
        if top < 0:
            top = 0
            bot = h
        if bot > maxbot:
            bot = maxbot
            top = bot - h

        nrect = SMALL_RECT()
        nrect.Top = top
        nrect.Bottom = bot
        nrect.Left = rect.Left
        nrect.Right = rect.Right
        log("sn: top=%d bot=%d" % (top, bot))
        r = self.SetConsoleWindowInfo(self.hout, True, byref(nrect))
        log("r=%d" % r)

    def get(self):
        """Get next event from queue."""
        inputHookFunc = c_void_p.from_address(self.inputHookPtr).value

        Cevent = INPUT_RECORD()
        count = DWORD(0)
        while True:
            if inputHookFunc:
                call_function(inputHookFunc, ())
            status = self.ReadConsoleInputW(self.hin, byref(Cevent), 1, byref(count))
            if status and count.value == 1:
                e = event(self, Cevent)
                return e

    def getkeypress(self):
        """Return next key press event from the queue, ignoring others."""
        while True:
            e = self.get()
            if e.type == "KeyPress" and e.keycode not in key_modifiers:
                log("console.getkeypress %s" % e)
                if e.keyinfo.keyname == "next":
                    self.scroll_window(12)
                elif e.keyinfo.keyname == "prior":
                    self.scroll_window(-12)
                else:
                    return e
            elif (e.type == "KeyRelease") and (
                e.keyinfo == KeyPress("S", False, True, False, "S")
                or e.keyinfo == KeyPress("C", False, True, False, "C")
            ):
                log("getKeypress:%s,%s,%s" % (e.keyinfo, e.keycode, e.type))
                return e

    def getchar(self):
        """Get next character from queue."""

        Cevent = INPUT_RECORD()
        count = DWORD(0)
        while True:
            status = self.ReadConsoleInputW(self.hin, byref(Cevent), 1, byref(count))
            if (
                status
                and (count.value == 1)
                and (Cevent.EventType == 1)
                and Cevent.Event.KeyEvent.bKeyDown
            ):
                sym = keysym(Cevent.Event.KeyEvent.wVirtualKeyCode)
                if len(sym) == 0:
                    sym = Cevent.Event.KeyEvent.uChar.AsciiChar
                return sym

    def peek(self):
        """Check event queue."""
        Cevent = INPUT_RECORD()
        count = DWORD(0)
        status = self.PeekConsoleInputW(self.hin, byref(Cevent), 1, byref(count))
        if status and count == 1:
            return event(self, Cevent)

    def title(self, txt=None):
        """Set/get title."""
        if txt:
            self.SetConsoleTitleW(txt)
        else:
            buffer = create_unicode_buffer(200)
            n = self.GetConsoleTitleW(buffer, 200)
            if n > 0:
                return buffer.value[:n]

    def size(self, width=None, height=None):
        """Set/get window size."""
        info = CONSOLE_SCREEN_BUFFER_INFO()
        status = self.GetConsoleScreenBufferInfo(self.hout, byref(info))
        if not status:
            return None
        if width is not None and height is not None:
            wmin = info.srWindow.Right - info.srWindow.Left + 1
            hmin = info.srWindow.Bottom - info.srWindow.Top + 1
            # print wmin, hmin
            width = max(width, wmin)
            height = max(height, hmin)
            # print width, height
            self.SetConsoleScreenBufferSize(self.hout, self.fixcoord(width, height))
        else:
            return (info.dwSize.X, info.dwSize.Y)

    def cursor(self, visible=None, size=None):
        """Set cursor on or off."""
        info = CONSOLE_CURSOR_INFO()
        if self.GetConsoleCursorInfo(self.hout, byref(info)):
            if visible is not None:
                info.bVisible = visible
            if size is not None:
                info.dwSize = size
            self.SetConsoleCursorInfo(self.hout, byref(info))

    def bell(self):
        self.write("\007")

    def next_serial(self):
        """Get next event serial number."""
        self.serial += 1
        return self.serial


# add the functions from the dll to the class
for func in funcs:
    setattr(Console, func, getattr(windll.kernel32, func))


_strncpy = ctypes.windll.kernel32.lstrcpynA
_strncpy.restype = c_char_p
_strncpy.argtypes = [c_char_p, c_char_p, c_size_t]


LPVOID = c_void_p
LPCVOID = c_void_p
FARPROC = c_void_p
LPDWORD = POINTER(DWORD)

Console.AllocConsole.restype = BOOL
Console.AllocConsole.argtypes = []  # void
Console.CreateConsoleScreenBuffer.restype = HANDLE
Console.CreateConsoleScreenBuffer.argtypes = [
    DWORD,
    DWORD,
    c_void_p,
    DWORD,
    LPVOID,
]  # DWORD, DWORD, SECURITY_ATTRIBUTES*, DWORD, LPVOID
Console.FillConsoleOutputAttribute.restype = BOOL
Console.FillConsoleOutputAttribute.argtypes = [
    HANDLE,
    WORD,
    DWORD,
    c_int,
    LPDWORD,
]  # HANDLE, WORD, DWORD, COORD, LPDWORD
Console.FillConsoleOutputCharacterW.restype = BOOL
Console.FillConsoleOutputCharacterW.argtypes = [
    HANDLE,
    c_ushort,
    DWORD,
    c_int,
    LPDWORD,
]  # HANDLE, TCHAR, DWORD, COORD, LPDWORD
Console.FreeConsole.restype = BOOL
Console.FreeConsole.argtypes = []  # void
Console.GetConsoleCursorInfo.restype = BOOL
Console.GetConsoleCursorInfo.argtypes = [
    HANDLE,
    c_void_p,
]  # HANDLE, PCONSOLE_CURSOR_INFO
Console.GetConsoleMode.restype = BOOL
Console.GetConsoleMode.argtypes = [HANDLE, LPDWORD]  # HANDLE, LPDWORD
Console.GetConsoleScreenBufferInfo.restype = BOOL
Console.GetConsoleScreenBufferInfo.argtypes = [
    HANDLE,
    c_void_p,
]  # HANDLE, PCONSOLE_SCREEN_BUFFER_INFO
Console.GetConsoleTitleW.restype = DWORD
Console.GetConsoleTitleW.argtypes = [c_wchar_p, DWORD]  # LPTSTR , DWORD
Console.GetProcAddress.restype = FARPROC
Console.GetProcAddress.argtypes = [HMODULE, c_char_p]  # HMODULE , LPCSTR
Console.GetStdHandle.restype = HANDLE
Console.GetStdHandle.argtypes = [DWORD]
Console.PeekConsoleInputW.restype = BOOL
# HANDLE, PINPUT_RECORD, DWORD, LPDWORD
Console.PeekConsoleInputW.argtypes = [HANDLE, c_void_p, DWORD, LPDWORD]
Console.ReadConsoleInputW.restype = BOOL
# HANDLE, PINPUT_RECORD, DWORD, LPDWORD
Console.ReadConsoleInputW.argtypes = [HANDLE, c_void_p, DWORD, LPDWORD]
Console.ScrollConsoleScreenBufferW.restype = BOOL
Console.ScrollConsoleScreenBufferW.argtypes = [
    HANDLE,
    c_void_p,
    c_void_p,
    c_int,
    c_void_p,
]  # HANDLE, SMALL_RECT*, SMALL_RECT*, COORD, LPDWORD
Console.SetConsoleActiveScreenBuffer.restype = BOOL
Console.SetConsoleActiveScreenBuffer.argtypes = [HANDLE]  # HANDLE
Console.SetConsoleCursorInfo.restype = BOOL
Console.SetConsoleCursorInfo.argtypes = [
    HANDLE,
    c_void_p,
]  # HANDLE, CONSOLE_CURSOR_INFO*
Console.SetConsoleCursorPosition.restype = BOOL
Console.SetConsoleCursorPosition.argtypes = [HANDLE, c_int]  # HANDLE, COORD
Console.SetConsoleMode.restype = BOOL
Console.SetConsoleMode.argtypes = [HANDLE, DWORD]  # HANDLE, DWORD
Console.SetConsoleScreenBufferSize.restype = BOOL
Console.SetConsoleScreenBufferSize.argtypes = [HANDLE, c_int]  # HANDLE, COORD
Console.SetConsoleTextAttribute.restype = BOOL
Console.SetConsoleTextAttribute.argtypes = [HANDLE, WORD]  # HANDLE, WORD
Console.SetConsoleTitleW.restype = BOOL
Console.SetConsoleTitleW.argtypes = [c_wchar_p]  # LPCTSTR
Console.SetConsoleWindowInfo.restype = BOOL
Console.SetConsoleWindowInfo.argtypes = [
    HANDLE,
    BOOL,
    c_void_p,
]  # HANDLE, BOOL, SMALL_RECT*
Console.WriteConsoleW.restype = BOOL
# HANDLE, VOID*, DWORD, LPDWORD, LPVOID
Console.WriteConsoleW.argtypes = [HANDLE, c_void_p, DWORD, LPDWORD, LPVOID]
Console.WriteConsoleOutputCharacterW.restype = BOOL
Console.WriteConsoleOutputCharacterW.argtypes = [
    HANDLE,
    c_wchar_p,
    DWORD,
    c_int,
    LPDWORD,
]  # HANDLE, LPCTSTR, DWORD, COORD, LPDWORD
Console.WriteFile.restype = BOOL
# HANDLE, LPCVOID , DWORD, LPDWORD , LPOVERLAPPED
Console.WriteFile.argtypes = [HANDLE, LPCVOID, DWORD, LPDWORD, c_void_p]


VkKeyScan = windll.user32.VkKeyScanA


class event(Event):
    """Represent events from the console."""

    def __init__(self, console, input):
        """Initialize an event from the Windows input structure."""
        self.type = "??"
        self.serial = console.next_serial()
        self.width = 0
        self.height = 0
        self.x = 0
        self.y = 0
        self.char = ""
        self.keycode = 0
        self.keysym = "??"
        # a tuple with (control, meta, shift, keycode) for dispatch
        self.keyinfo = None
        self.width = None

        if input.EventType == KEY_EVENT:
            if input.Event.KeyEvent.bKeyDown:
                self.type = "KeyPress"
            else:
                self.type = "KeyRelease"
            self.char = input.Event.KeyEvent.uChar.UnicodeChar
            self.keycode = input.Event.KeyEvent.wVirtualKeyCode
            self.state = input.Event.KeyEvent.dwControlKeyState
            self.keyinfo = make_KeyPress(self.char, self.state, self.keycode)

        elif input.EventType == MOUSE_EVENT:
            if input.Event.MouseEvent.dwEventFlags & MOUSE_MOVED:
                self.type = "Motion"
            else:
                self.type = "Button"
            self.x = input.Event.MouseEvent.dwMousePosition.X
            self.y = input.Event.MouseEvent.dwMousePosition.Y
            self.state = input.Event.MouseEvent.dwButtonState
        elif input.EventType == WINDOW_BUFFER_SIZE_EVENT:
            self.type = "Configure"
            self.width = input.Event.WindowBufferSizeEvent.dwSize.X
            self.height = input.Event.WindowBufferSizeEvent.dwSize.Y
        elif input.EventType == FOCUS_EVENT:
            if input.Event.FocusEvent.bSetFocus:
                self.type = "FocusIn"
            else:
                self.type = "FocusOut"
        elif input.EventType == MENU_EVENT:
            self.type = "Menu"
            self.state = input.Event.MenuEvent.dwCommandId


def getconsole(buffer=1):
    """Get a console handle.

    If buffer is non-zero, a new console buffer is allocated and
    installed.  Otherwise, this returns a handle to the current
    console buffer"""

    c = Console(buffer)

    return c


# The following code uses ctypes to allow a Python callable to
# substitute for GNU readline within the Python interpreter. Calling
# raw_input or other functions that do input, inside your callable
# might be a bad idea, then again, it might work.

# The Python callable can raise EOFError or KeyboardInterrupt and
# these will be translated into the appropriate outputs from readline
# so that they will then be translated back!

# If the Python callable raises any other exception, a traceback will
# be printed and readline will appear to return an empty line.

# I use ctypes to create a C-callable from a Python wrapper that
# handles the exceptions and gets the result into the right form.


# the type for our C-callable wrapper
HOOKFUNC23 = CFUNCTYPE(c_char_p, c_void_p, c_void_p, c_char_p)

readline_hook = None  # the python hook goes here
readline_ref = None  # reference to the c-callable to keep it alive


def hook_wrapper_23(stdin, stdout, prompt):
    """Wrap a Python readline so it behaves like GNU readline."""
    try:
        # call the Python hook
        res = ensure_str(readline_hook(prompt))
        # make sure it returned the right sort of thing
        if res and not isinstance(res, bytes):
            raise TypeError("readline must return a string.")
    except KeyboardInterrupt:
        # GNU readline returns 0 on keyboard interrupt
        return 0
    except EOFError:
        # It returns an empty string on EOF
        res = ensure_str("")
    except BaseException:
        print("Readline internal error", file=sys.stderr)
        traceback.print_exc()
        res = ensure_str("\n")
    # we have to make a copy because the caller expects to free the result
    n = len(res)
    p = Console.PyMem_Malloc(n + 1)
    _strncpy(cast(p, c_char_p), res, n + 1)
    return p


def install_readline(hook):
    """Set up things for the interpreter to call
    our function like GNU readline."""
    global readline_hook, readline_ref
    # save the hook so the wrapper can call it
    readline_hook = hook
    # get the address of PyOS_ReadlineFunctionPointer so we can update it
    PyOS_RFP = c_void_p.from_address(
        Console.GetProcAddress(
            sys.dllhandle, "PyOS_ReadlineFunctionPointer".encode("ascii")
        )
    )
    # save a reference to the generated C-callable so it doesn't go away
    readline_ref = HOOKFUNC23(hook_wrapper_23)
    # get the address of the function
    func_start = c_void_p.from_address(addressof(readline_ref)).value
    # write the function address into PyOS_ReadlineFunctionPointer
    PyOS_RFP.value = func_start


if __name__ == "__main__":
    import sys
    import time

    def p(char):
        return chr(VkKeyScan(ord(char)) & 0xFF)

    c = Console(0)
    sys.stdout = c
    sys.stderr = c
    c.page()
    print(p("d"), p("D"))
    c.pos(5, 10)
    c.write("hi there")
    print("some printed output")
    for i in range(10):
        q = c.getkeypress()
        print(q)
    del c

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\reshape\pivot.py ===
from __future__ import annotations

from collections.abc import (
    Hashable,
    Sequence,
)
from typing import (
    TYPE_CHECKING,
    Callable,
    Literal,
    cast,
)
import warnings

import numpy as np

from pandas._libs import lib
from pandas.util._decorators import (
    Appender,
    Substitution,
)
from pandas.util._exceptions import find_stack_level

from pandas.core.dtypes.cast import maybe_downcast_to_dtype
from pandas.core.dtypes.common import (
    is_list_like,
    is_nested_list_like,
    is_scalar,
)
from pandas.core.dtypes.dtypes import ExtensionDtype
from pandas.core.dtypes.generic import (
    ABCDataFrame,
    ABCSeries,
)

import pandas.core.common as com
from pandas.core.frame import _shared_docs
from pandas.core.groupby import Grouper
from pandas.core.indexes.api import (
    Index,
    MultiIndex,
    get_objs_combined_axis,
)
from pandas.core.reshape.concat import concat
from pandas.core.reshape.util import cartesian_product
from pandas.core.series import Series

if TYPE_CHECKING:
    from pandas._typing import (
        AggFuncType,
        AggFuncTypeBase,
        AggFuncTypeDict,
        IndexLabel,
    )

    from pandas import DataFrame


# Note: We need to make sure `frame` is imported before `pivot`, otherwise
# _shared_docs['pivot_table'] will not yet exist.  TODO: Fix this dependency
@Substitution("\ndata : DataFrame")
@Appender(_shared_docs["pivot_table"], indents=1)
def pivot_table(
    data: DataFrame,
    values=None,
    index=None,
    columns=None,
    aggfunc: AggFuncType = "mean",
    fill_value=None,
    margins: bool = False,
    dropna: bool = True,
    margins_name: Hashable = "All",
    observed: bool | lib.NoDefault = lib.no_default,
    sort: bool = True,
) -> DataFrame:
    index = _convert_by(index)
    columns = _convert_by(columns)

    if isinstance(aggfunc, list):
        pieces: list[DataFrame] = []
        keys = []
        for func in aggfunc:
            _table = __internal_pivot_table(
                data,
                values=values,
                index=index,
                columns=columns,
                fill_value=fill_value,
                aggfunc=func,
                margins=margins,
                dropna=dropna,
                margins_name=margins_name,
                observed=observed,
                sort=sort,
            )
            pieces.append(_table)
            keys.append(getattr(func, "__name__", func))

        table = concat(pieces, keys=keys, axis=1)
        return table.__finalize__(data, method="pivot_table")

    table = __internal_pivot_table(
        data,
        values,
        index,
        columns,
        aggfunc,
        fill_value,
        margins,
        dropna,
        margins_name,
        observed,
        sort,
    )
    return table.__finalize__(data, method="pivot_table")


def __internal_pivot_table(
    data: DataFrame,
    values,
    index,
    columns,
    aggfunc: AggFuncTypeBase | AggFuncTypeDict,
    fill_value,
    margins: bool,
    dropna: bool,
    margins_name: Hashable,
    observed: bool | lib.NoDefault,
    sort: bool,
) -> DataFrame:
    """
    Helper of :func:`pandas.pivot_table` for any non-list ``aggfunc``.
    """
    keys = index + columns

    values_passed = values is not None
    if values_passed:
        if is_list_like(values):
            values_multi = True
            values = list(values)
        else:
            values_multi = False
            values = [values]

        # GH14938 Make sure value labels are in data
        for i in values:
            if i not in data:
                raise KeyError(i)

        to_filter = []
        for x in keys + values:
            if isinstance(x, Grouper):
                x = x.key
            try:
                if x in data:
                    to_filter.append(x)
            except TypeError:
                pass
        if len(to_filter) < len(data.columns):
            data = data[to_filter]

    else:
        values = data.columns
        for key in keys:
            try:
                values = values.drop(key)
            except (TypeError, ValueError, KeyError):
                pass
        values = list(values)

    observed_bool = False if observed is lib.no_default else observed
    grouped = data.groupby(keys, observed=observed_bool, sort=sort, dropna=dropna)
    if observed is lib.no_default and any(
        ping._passed_categorical for ping in grouped._grouper.groupings
    ):
        warnings.warn(
            "The default value of observed=False is deprecated and will change "
            "to observed=True in a future version of pandas. Specify "
            "observed=False to silence this warning and retain the current behavior",
            category=FutureWarning,
            stacklevel=find_stack_level(),
        )
    agged = grouped.agg(aggfunc)

    if dropna and isinstance(agged, ABCDataFrame) and len(agged.columns):
        agged = agged.dropna(how="all")

    table = agged

    # GH17038, this check should only happen if index is defined (not None)
    if table.index.nlevels > 1 and index:
        # Related GH #17123
        # If index_names are integers, determine whether the integers refer
        # to the level position or name.
        index_names = agged.index.names[: len(index)]
        to_unstack = []
        for i in range(len(index), len(keys)):
            name = agged.index.names[i]
            if name is None or name in index_names:
                to_unstack.append(i)
            else:
                to_unstack.append(name)
        table = agged.unstack(to_unstack, fill_value=fill_value)

    if not dropna:
        if isinstance(table.index, MultiIndex):
            m = MultiIndex.from_arrays(
                cartesian_product(table.index.levels), names=table.index.names
            )
            table = table.reindex(m, axis=0, fill_value=fill_value)

        if isinstance(table.columns, MultiIndex):
            m = MultiIndex.from_arrays(
                cartesian_product(table.columns.levels), names=table.columns.names
            )
            table = table.reindex(m, axis=1, fill_value=fill_value)

    if sort is True and isinstance(table, ABCDataFrame):
        table = table.sort_index(axis=1)

    if fill_value is not None:
        table = table.fillna(fill_value)
        if aggfunc is len and not observed and lib.is_integer(fill_value):
            # TODO: can we avoid this?  this used to be handled by
            #  downcast="infer" in fillna
            table = table.astype(np.int64)

    if margins:
        if dropna:
            data = data[data.notna().all(axis=1)]
        table = _add_margins(
            table,
            data,
            values,
            rows=index,
            cols=columns,
            aggfunc=aggfunc,
            observed=dropna,
            margins_name=margins_name,
            fill_value=fill_value,
        )

    # discard the top level
    if values_passed and not values_multi and table.columns.nlevels > 1:
        table.columns = table.columns.droplevel(0)
    if len(index) == 0 and len(columns) > 0:
        table = table.T

    # GH 15193 Make sure empty columns are removed if dropna=True
    if isinstance(table, ABCDataFrame) and dropna:
        table = table.dropna(how="all", axis=1)

    return table


def _add_margins(
    table: DataFrame | Series,
    data: DataFrame,
    values,
    rows,
    cols,
    aggfunc,
    observed: bool,
    margins_name: Hashable = "All",
    fill_value=None,
):
    if not isinstance(margins_name, str):
        raise ValueError("margins_name argument must be a string")

    msg = f'Conflicting name "{margins_name}" in margins'
    for level in table.index.names:
        if margins_name in table.index.get_level_values(level):
            raise ValueError(msg)

    grand_margin = _compute_grand_margin(data, values, aggfunc, margins_name)

    if table.ndim == 2:
        # i.e. DataFrame
        for level in table.columns.names[1:]:
            if margins_name in table.columns.get_level_values(level):
                raise ValueError(msg)

    key: str | tuple[str, ...]
    if len(rows) > 1:
        key = (margins_name,) + ("",) * (len(rows) - 1)
    else:
        key = margins_name

    if not values and isinstance(table, ABCSeries):
        # If there are no values and the table is a series, then there is only
        # one column in the data. Compute grand margin and return it.
        return table._append(table._constructor({key: grand_margin[margins_name]}))

    elif values:
        marginal_result_set = _generate_marginal_results(
            table, data, values, rows, cols, aggfunc, observed, margins_name
        )
        if not isinstance(marginal_result_set, tuple):
            return marginal_result_set
        result, margin_keys, row_margin = marginal_result_set
    else:
        # no values, and table is a DataFrame
        assert isinstance(table, ABCDataFrame)
        marginal_result_set = _generate_marginal_results_without_values(
            table, data, rows, cols, aggfunc, observed, margins_name
        )
        if not isinstance(marginal_result_set, tuple):
            return marginal_result_set
        result, margin_keys, row_margin = marginal_result_set

    row_margin = row_margin.reindex(result.columns, fill_value=fill_value)
    # populate grand margin
    for k in margin_keys:
        if isinstance(k, str):
            row_margin[k] = grand_margin[k]
        else:
            row_margin[k] = grand_margin[k[0]]

    from pandas import DataFrame

    margin_dummy = DataFrame(row_margin, columns=Index([key])).T

    row_names = result.index.names
    # check the result column and leave floats

    for dtype in set(result.dtypes):
        if isinstance(dtype, ExtensionDtype):
            # Can hold NA already
            continue

        cols = result.select_dtypes([dtype]).columns
        margin_dummy[cols] = margin_dummy[cols].apply(
            maybe_downcast_to_dtype, args=(dtype,)
        )
    result = result._append(margin_dummy)
    result.index.names = row_names

    return result


def _compute_grand_margin(
    data: DataFrame, values, aggfunc, margins_name: Hashable = "All"
):
    if values:
        grand_margin = {}
        for k, v in data[values].items():
            try:
                if isinstance(aggfunc, str):
                    grand_margin[k] = getattr(v, aggfunc)()
                elif isinstance(aggfunc, dict):
                    if isinstance(aggfunc[k], str):
                        grand_margin[k] = getattr(v, aggfunc[k])()
                    else:
                        grand_margin[k] = aggfunc[k](v)
                else:
                    grand_margin[k] = aggfunc(v)
            except TypeError:
                pass
        return grand_margin
    else:
        return {margins_name: aggfunc(data.index)}


def _generate_marginal_results(
    table,
    data: DataFrame,
    values,
    rows,
    cols,
    aggfunc,
    observed: bool,
    margins_name: Hashable = "All",
):
    margin_keys: list | Index
    if len(cols) > 0:
        # need to "interleave" the margins
        table_pieces = []
        margin_keys = []

        def _all_key(key):
            return (key, margins_name) + ("",) * (len(cols) - 1)

        if len(rows) > 0:
            margin = data[rows + values].groupby(rows, observed=observed).agg(aggfunc)
            cat_axis = 1

            for key, piece in table.T.groupby(level=0, observed=observed):
                piece = piece.T
                all_key = _all_key(key)

                # we are going to mutate this, so need to copy!
                piece = piece.copy()
                piece[all_key] = margin[key]

                table_pieces.append(piece)
                margin_keys.append(all_key)
        else:
            from pandas import DataFrame

            cat_axis = 0
            for key, piece in table.groupby(level=0, observed=observed):
                if len(cols) > 1:
                    all_key = _all_key(key)
                else:
                    all_key = margins_name
                table_pieces.append(piece)
                # GH31016 this is to calculate margin for each group, and assign
                # corresponded key as index
                transformed_piece = DataFrame(piece.apply(aggfunc)).T
                if isinstance(piece.index, MultiIndex):
                    # We are adding an empty level
                    transformed_piece.index = MultiIndex.from_tuples(
                        [all_key], names=piece.index.names + [None]
                    )
                else:
                    transformed_piece.index = Index([all_key], name=piece.index.name)

                # append piece for margin into table_piece
                table_pieces.append(transformed_piece)
                margin_keys.append(all_key)

        if not table_pieces:
            # GH 49240
            return table
        else:
            result = concat(table_pieces, axis=cat_axis)

        if len(rows) == 0:
            return result
    else:
        result = table
        margin_keys = table.columns

    if len(cols) > 0:
        row_margin = data[cols + values].groupby(cols, observed=observed).agg(aggfunc)
        row_margin = row_margin.stack(future_stack=True)

        # GH#26568. Use names instead of indices in case of numeric names
        new_order_indices = [len(cols)] + list(range(len(cols)))
        new_order_names = [row_margin.index.names[i] for i in new_order_indices]
        row_margin.index = row_margin.index.reorder_levels(new_order_names)
    else:
        row_margin = data._constructor_sliced(np.nan, index=result.columns)

    return result, margin_keys, row_margin


def _generate_marginal_results_without_values(
    table: DataFrame,
    data: DataFrame,
    rows,
    cols,
    aggfunc,
    observed: bool,
    margins_name: Hashable = "All",
):
    margin_keys: list | Index
    if len(cols) > 0:
        # need to "interleave" the margins
        margin_keys = []

        def _all_key():
            if len(cols) == 1:
                return margins_name
            return (margins_name,) + ("",) * (len(cols) - 1)

        if len(rows) > 0:
            margin = data.groupby(rows, observed=observed)[rows].apply(aggfunc)
            all_key = _all_key()
            table[all_key] = margin
            result = table
            margin_keys.append(all_key)

        else:
            margin = data.groupby(level=0, axis=0, observed=observed).apply(aggfunc)
            all_key = _all_key()
            table[all_key] = margin
            result = table
            margin_keys.append(all_key)
            return result
    else:
        result = table
        margin_keys = table.columns

    if len(cols):
        row_margin = data.groupby(cols, observed=observed)[cols].apply(aggfunc)
    else:
        row_margin = Series(np.nan, index=result.columns)

    return result, margin_keys, row_margin


def _convert_by(by):
    if by is None:
        by = []
    elif (
        is_scalar(by)
        or isinstance(by, (np.ndarray, Index, ABCSeries, Grouper))
        or callable(by)
    ):
        by = [by]
    else:
        by = list(by)
    return by


@Substitution("\ndata : DataFrame")
@Appender(_shared_docs["pivot"], indents=1)
def pivot(
    data: DataFrame,
    *,
    columns: IndexLabel,
    index: IndexLabel | lib.NoDefault = lib.no_default,
    values: IndexLabel | lib.NoDefault = lib.no_default,
) -> DataFrame:
    columns_listlike = com.convert_to_list_like(columns)

    # If columns is None we will create a MultiIndex level with None as name
    # which might cause duplicated names because None is the default for
    # level names
    data = data.copy(deep=False)
    data.index = data.index.copy()
    data.index.names = [
        name if name is not None else lib.no_default for name in data.index.names
    ]

    indexed: DataFrame | Series
    if values is lib.no_default:
        if index is not lib.no_default:
            cols = com.convert_to_list_like(index)
        else:
            cols = []

        append = index is lib.no_default
        # error: Unsupported operand types for + ("List[Any]" and "ExtensionArray")
        # error: Unsupported left operand type for + ("ExtensionArray")
        indexed = data.set_index(
            cols + columns_listlike, append=append  # type: ignore[operator]
        )
    else:
        index_list: list[Index] | list[Series]
        if index is lib.no_default:
            if isinstance(data.index, MultiIndex):
                # GH 23955
                index_list = [
                    data.index.get_level_values(i) for i in range(data.index.nlevels)
                ]
            else:
                index_list = [
                    data._constructor_sliced(data.index, name=data.index.name)
                ]
        else:
            index_list = [data[idx] for idx in com.convert_to_list_like(index)]

        data_columns = [data[col] for col in columns_listlike]
        index_list.extend(data_columns)
        multiindex = MultiIndex.from_arrays(index_list)

        if is_list_like(values) and not isinstance(values, tuple):
            # Exclude tuple because it is seen as a single column name
            values = cast(Sequence[Hashable], values)
            indexed = data._constructor(
                data[values]._values, index=multiindex, columns=values
            )
        else:
            indexed = data._constructor_sliced(data[values]._values, index=multiindex)
    # error: Argument 1 to "unstack" of "DataFrame" has incompatible type "Union
    # [List[Any], ExtensionArray, ndarray[Any, Any], Index, Series]"; expected
    # "Hashable"
    result = indexed.unstack(columns_listlike)  # type: ignore[arg-type]
    result.index.names = [
        name if name is not lib.no_default else None for name in result.index.names
    ]

    return result


def crosstab(
    index,
    columns,
    values=None,
    rownames=None,
    colnames=None,
    aggfunc=None,
    margins: bool = False,
    margins_name: Hashable = "All",
    dropna: bool = True,
    normalize: bool | Literal[0, 1, "all", "index", "columns"] = False,
) -> DataFrame:
    """
    Compute a simple cross tabulation of two (or more) factors.

    By default, computes a frequency table of the factors unless an
    array of values and an aggregation function are passed.

    Parameters
    ----------
    index : array-like, Series, or list of arrays/Series
        Values to group by in the rows.
    columns : array-like, Series, or list of arrays/Series
        Values to group by in the columns.
    values : array-like, optional
        Array of values to aggregate according to the factors.
        Requires `aggfunc` be specified.
    rownames : sequence, default None
        If passed, must match number of row arrays passed.
    colnames : sequence, default None
        If passed, must match number of column arrays passed.
    aggfunc : function, optional
        If specified, requires `values` be specified as well.
    margins : bool, default False
        Add row/column margins (subtotals).
    margins_name : str, default 'All'
        Name of the row/column that will contain the totals
        when margins is True.
    dropna : bool, default True
        Do not include columns whose entries are all NaN.
    normalize : bool, {'all', 'index', 'columns'}, or {0,1}, default False
        Normalize by dividing all values by the sum of values.

        - If passed 'all' or `True`, will normalize over all values.
        - If passed 'index' will normalize over each row.
        - If passed 'columns' will normalize over each column.
        - If margins is `True`, will also normalize margin values.

    Returns
    -------
    DataFrame
        Cross tabulation of the data.

    See Also
    --------
    DataFrame.pivot : Reshape data based on column values.
    pivot_table : Create a pivot table as a DataFrame.

    Notes
    -----
    Any Series passed will have their name attributes used unless row or column
    names for the cross-tabulation are specified.

    Any input passed containing Categorical data will have **all** of its
    categories included in the cross-tabulation, even if the actual data does
    not contain any instances of a particular category.

    In the event that there aren't overlapping indexes an empty DataFrame will
    be returned.

    Reference :ref:`the user guide <reshaping.crosstabulations>` for more examples.

    Examples
    --------
    >>> a = np.array(["foo", "foo", "foo", "foo", "bar", "bar",
    ...               "bar", "bar", "foo", "foo", "foo"], dtype=object)
    >>> b = np.array(["one", "one", "one", "two", "one", "one",
    ...               "one", "two", "two", "two", "one"], dtype=object)
    >>> c = np.array(["dull", "dull", "shiny", "dull", "dull", "shiny",
    ...               "shiny", "dull", "shiny", "shiny", "shiny"],
    ...              dtype=object)
    >>> pd.crosstab(a, [b, c], rownames=['a'], colnames=['b', 'c'])
    b   one        two
    c   dull shiny dull shiny
    a
    bar    1     2    1     0
    foo    2     2    1     2

    Here 'c' and 'f' are not represented in the data and will not be
    shown in the output because dropna is True by default. Set
    dropna=False to preserve categories with no data.

    >>> foo = pd.Categorical(['a', 'b'], categories=['a', 'b', 'c'])
    >>> bar = pd.Categorical(['d', 'e'], categories=['d', 'e', 'f'])
    >>> pd.crosstab(foo, bar)
    col_0  d  e
    row_0
    a      1  0
    b      0  1
    >>> pd.crosstab(foo, bar, dropna=False)
    col_0  d  e  f
    row_0
    a      1  0  0
    b      0  1  0
    c      0  0  0
    """
    if values is None and aggfunc is not None:
        raise ValueError("aggfunc cannot be used without values.")

    if values is not None and aggfunc is None:
        raise ValueError("values cannot be used without an aggfunc.")

    if not is_nested_list_like(index):
        index = [index]
    if not is_nested_list_like(columns):
        columns = [columns]

    common_idx = None
    pass_objs = [x for x in index + columns if isinstance(x, (ABCSeries, ABCDataFrame))]
    if pass_objs:
        common_idx = get_objs_combined_axis(pass_objs, intersect=True, sort=False)

    rownames = _get_names(index, rownames, prefix="row")
    colnames = _get_names(columns, colnames, prefix="col")

    # duplicate names mapped to unique names for pivot op
    (
        rownames_mapper,
        unique_rownames,
        colnames_mapper,
        unique_colnames,
    ) = _build_names_mapper(rownames, colnames)

    from pandas import DataFrame

    data = {
        **dict(zip(unique_rownames, index)),
        **dict(zip(unique_colnames, columns)),
    }
    df = DataFrame(data, index=common_idx)

    if values is None:
        df["__dummy__"] = 0
        kwargs = {"aggfunc": len, "fill_value": 0}
    else:
        df["__dummy__"] = values
        kwargs = {"aggfunc": aggfunc}

    # error: Argument 7 to "pivot_table" of "DataFrame" has incompatible type
    # "**Dict[str, object]"; expected "Union[...]"
    table = df.pivot_table(
        "__dummy__",
        index=unique_rownames,
        columns=unique_colnames,
        margins=margins,
        margins_name=margins_name,
        dropna=dropna,
        observed=False,
        **kwargs,  # type: ignore[arg-type]
    )

    # Post-process
    if normalize is not False:
        table = _normalize(
            table, normalize=normalize, margins=margins, margins_name=margins_name
        )

    table = table.rename_axis(index=rownames_mapper, axis=0)
    table = table.rename_axis(columns=colnames_mapper, axis=1)

    return table


def _normalize(
    table: DataFrame, normalize, margins: bool, margins_name: Hashable = "All"
) -> DataFrame:
    if not isinstance(normalize, (bool, str)):
        axis_subs = {0: "index", 1: "columns"}
        try:
            normalize = axis_subs[normalize]
        except KeyError as err:
            raise ValueError("Not a valid normalize argument") from err

    if margins is False:
        # Actual Normalizations
        normalizers: dict[bool | str, Callable] = {
            "all": lambda x: x / x.sum(axis=1).sum(axis=0),
            "columns": lambda x: x / x.sum(),
            "index": lambda x: x.div(x.sum(axis=1), axis=0),
        }

        normalizers[True] = normalizers["all"]

        try:
            f = normalizers[normalize]
        except KeyError as err:
            raise ValueError("Not a valid normalize argument") from err

        table = f(table)
        table = table.fillna(0)

    elif margins is True:
        # keep index and column of pivoted table
        table_index = table.index
        table_columns = table.columns
        last_ind_or_col = table.iloc[-1, :].name

        # check if margin name is not in (for MI cases) and not equal to last
        # index/column and save the column and index margin
        if (margins_name not in last_ind_or_col) & (margins_name != last_ind_or_col):
            raise ValueError(f"{margins_name} not in pivoted DataFrame")
        column_margin = table.iloc[:-1, -1]
        index_margin = table.iloc[-1, :-1]

        # keep the core table
        table = table.iloc[:-1, :-1]

        # Normalize core
        table = _normalize(table, normalize=normalize, margins=False)

        # Fix Margins
        if normalize == "columns":
            column_margin = column_margin / column_margin.sum()
            table = concat([table, column_margin], axis=1)
            table = table.fillna(0)
            table.columns = table_columns

        elif normalize == "index":
            index_margin = index_margin / index_margin.sum()
            table = table._append(index_margin)
            table = table.fillna(0)
            table.index = table_index

        elif normalize == "all" or normalize is True:
            column_margin = column_margin / column_margin.sum()
            index_margin = index_margin / index_margin.sum()
            index_margin.loc[margins_name] = 1
            table = concat([table, column_margin], axis=1)
            table = table._append(index_margin)

            table = table.fillna(0)
            table.index = table_index
            table.columns = table_columns

        else:
            raise ValueError("Not a valid normalize argument")

    else:
        raise ValueError("Not a valid margins argument")

    return table


def _get_names(arrs, names, prefix: str = "row"):
    if names is None:
        names = []
        for i, arr in enumerate(arrs):
            if isinstance(arr, ABCSeries) and arr.name is not None:
                names.append(arr.name)
            else:
                names.append(f"{prefix}_{i}")
    else:
        if len(names) != len(arrs):
            raise AssertionError("arrays and names must have the same length")
        if not isinstance(names, list):
            names = list(names)

    return names


def _build_names_mapper(
    rownames: list[str], colnames: list[str]
) -> tuple[dict[str, str], list[str], dict[str, str], list[str]]:
    """
    Given the names of a DataFrame's rows and columns, returns a set of unique row
    and column names and mappers that convert to original names.

    A row or column name is replaced if it is duplicate among the rows of the inputs,
    among the columns of the inputs or between the rows and the columns.

    Parameters
    ----------
    rownames: list[str]
    colnames: list[str]

    Returns
    -------
    Tuple(Dict[str, str], List[str], Dict[str, str], List[str])

    rownames_mapper: dict[str, str]
        a dictionary with new row names as keys and original rownames as values
    unique_rownames: list[str]
        a list of rownames with duplicate names replaced by dummy names
    colnames_mapper: dict[str, str]
        a dictionary with new column names as keys and original column names as values
    unique_colnames: list[str]
        a list of column names with duplicate names replaced by dummy names

    """

    def get_duplicates(names):
        seen: set = set()
        return {name for name in names if name not in seen}

    shared_names = set(rownames).intersection(set(colnames))
    dup_names = get_duplicates(rownames) | get_duplicates(colnames) | shared_names

    rownames_mapper = {
        f"row_{i}": name for i, name in enumerate(rownames) if name in dup_names
    }
    unique_rownames = [
        f"row_{i}" if name in dup_names else name for i, name in enumerate(rownames)
    ]

    colnames_mapper = {
        f"col_{i}": name for i, name in enumerate(colnames) if name in dup_names
    }
    unique_colnames = [
        f"col_{i}" if name in dup_names else name for i, name in enumerate(colnames)
    ]

    return rownames_mapper, unique_rownames, colnames_mapper, unique_colnames

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\calculus\util.py ===
from .accumulationbounds import AccumBounds, AccumulationBounds # noqa: F401
from .singularities import singularities
from sympy.core import Pow, S
from sympy.core.function import diff, expand_mul, Function
from sympy.core.kind import NumberKind
from sympy.core.mod import Mod
from sympy.core.numbers import equal_valued
from sympy.core.relational import Relational
from sympy.core.symbol import Symbol, Dummy
from sympy.core.sympify import _sympify
from sympy.functions.elementary.complexes import Abs, im, re
from sympy.functions.elementary.exponential import exp, log
from sympy.functions.elementary.integers import frac
from sympy.functions.elementary.piecewise import Piecewise
from sympy.functions.elementary.trigonometric import (
    TrigonometricFunction, sin, cos, tan, cot, csc, sec,
    asin, acos, acot, atan, asec, acsc)
from sympy.functions.elementary.hyperbolic import (sinh, cosh, tanh, coth,
    sech, csch, asinh, acosh, atanh, acoth, asech, acsch)
from sympy.polys.polytools import degree, lcm_list
from sympy.sets.sets import (Interval, Intersection, FiniteSet, Union,
                             Complement)
from sympy.sets.fancysets import ImageSet
from sympy.sets.conditionset import ConditionSet
from sympy.utilities import filldedent
from sympy.utilities.iterables import iterable
from sympy.matrices.dense import hessian


def continuous_domain(f, symbol, domain):
    """
    Returns the domain on which the function expression f is continuous.

    This function is limited by the ability to determine the various
    singularities and discontinuities of the given function.
    The result is either given as a union of intervals or constructed using
    other set operations.

    Parameters
    ==========

    f : :py:class:`~.Expr`
        The concerned function.
    symbol : :py:class:`~.Symbol`
        The variable for which the intervals are to be determined.
    domain : :py:class:`~.Interval`
        The domain over which the continuity of the symbol has to be checked.

    Examples
    ========

    >>> from sympy import Interval, Symbol, S, tan, log, pi, sqrt
    >>> from sympy.calculus.util import continuous_domain
    >>> x = Symbol('x')
    >>> continuous_domain(1/x, x, S.Reals)
    Union(Interval.open(-oo, 0), Interval.open(0, oo))
    >>> continuous_domain(tan(x), x, Interval(0, pi))
    Union(Interval.Ropen(0, pi/2), Interval.Lopen(pi/2, pi))
    >>> continuous_domain(sqrt(x - 2), x, Interval(-5, 5))
    Interval(2, 5)
    >>> continuous_domain(log(2*x - 1), x, S.Reals)
    Interval.open(1/2, oo)

    Returns
    =======

    :py:class:`~.Interval`
        Union of all intervals where the function is continuous.

    Raises
    ======

    NotImplementedError
        If the method to determine continuity of such a function
        has not yet been developed.

    """
    from sympy.solvers.inequalities import solve_univariate_inequality

    if not domain.is_subset(S.Reals):
        raise NotImplementedError(filldedent('''
            Domain must be a subset of S.Reals.
            '''))
    implemented = [Pow, exp, log, Abs, frac,
                   sin, cos, tan, cot, sec, csc,
                   asin, acos, atan, acot, asec, acsc,
                   sinh, cosh, tanh, coth, sech, csch,
                   asinh, acosh, atanh, acoth, asech, acsch]
    used = [fct.func for fct in f.atoms(Function) if fct.has(symbol)]
    if any(func not in implemented for func in used):
        raise NotImplementedError(filldedent('''
            Unable to determine the domain of the given function.
            '''))

    x = Symbol('x')
    constraints = {
        log: (x > 0,),
        asin: (x >= -1, x <= 1),
        acos: (x >= -1, x <= 1),
        acosh: (x >= 1,),
        atanh: (x > -1, x < 1),
        asech: (x > 0, x <= 1)
    }
    constraints_union = {
        asec: (x <= -1, x >= 1),
        acsc: (x <= -1, x >= 1),
        acoth: (x < -1, x > 1)
    }

    cont_domain = domain
    for atom in f.atoms(Pow):
        den = atom.exp.as_numer_denom()[1]
        if atom.exp.is_rational and den.is_odd:
            pass    # 0**negative handled by singularities()
        else:
            constraint = solve_univariate_inequality(atom.base >= 0,
                                                        symbol).as_set()
            cont_domain = Intersection(constraint, cont_domain)

    for atom in f.atoms(Function):
        if atom.func in constraints:
            for c in constraints[atom.func]:
                constraint_relational = c.subs(x, atom.args[0])
                constraint_set = solve_univariate_inequality(
                    constraint_relational, symbol).as_set()
                cont_domain = Intersection(constraint_set, cont_domain)
        elif atom.func in constraints_union:
            constraint_set = S.EmptySet
            for c in constraints_union[atom.func]:
                constraint_relational = c.subs(x, atom.args[0])
                constraint_set += solve_univariate_inequality(
                    constraint_relational, symbol).as_set()
            cont_domain = Intersection(constraint_set, cont_domain)
        # XXX: the discontinuities below could be factored out in
        # a new "discontinuities()".
        elif atom.func == acot:
            from sympy.solvers.solveset import solveset_real
            # Sympy's acot() has a step discontinuity at 0. Since it's
            # neither an essential singularity nor a pole, singularities()
            # will not report it. But it's still relevant for determining
            # the continuity of the function f.
            cont_domain -= solveset_real(atom.args[0], symbol)
            # Note that the above may introduce spurious discontinuities, e.g.
            # for abs(acot(x)) at 0.
        elif atom.func == frac:
            from sympy.solvers.solveset import solveset_real
            r = function_range(atom.args[0], symbol, domain)
            r = Intersection(r, S.Integers)
            if r.is_finite_set:
                discont = S.EmptySet
                for n in r:
                    discont += solveset_real(atom.args[0]-n, symbol)
            else:
                discont = ConditionSet(
                    symbol, S.Integers.contains(atom.args[0]), cont_domain)
            cont_domain -= discont

    return cont_domain - singularities(f, symbol, domain)


def function_range(f, symbol, domain):
    """
    Finds the range of a function in a given domain.
    This method is limited by the ability to determine the singularities and
    determine limits.

    Parameters
    ==========

    f : :py:class:`~.Expr`
        The concerned function.
    symbol : :py:class:`~.Symbol`
        The variable for which the range of function is to be determined.
    domain : :py:class:`~.Interval`
        The domain under which the range of the function has to be found.

    Examples
    ========

    >>> from sympy import Interval, Symbol, S, exp, log, pi, sqrt, sin, tan
    >>> from sympy.calculus.util import function_range
    >>> x = Symbol('x')
    >>> function_range(sin(x), x, Interval(0, 2*pi))
    Interval(-1, 1)
    >>> function_range(tan(x), x, Interval(-pi/2, pi/2))
    Interval(-oo, oo)
    >>> function_range(1/x, x, S.Reals)
    Union(Interval.open(-oo, 0), Interval.open(0, oo))
    >>> function_range(exp(x), x, S.Reals)
    Interval.open(0, oo)
    >>> function_range(log(x), x, S.Reals)
    Interval(-oo, oo)
    >>> function_range(sqrt(x), x, Interval(-5, 9))
    Interval(0, 3)

    Returns
    =======

    :py:class:`~.Interval`
        Union of all ranges for all intervals under domain where function is
        continuous.

    Raises
    ======

    NotImplementedError
        If any of the intervals, in the given domain, for which function
        is continuous are not finite or real,
        OR if the critical points of the function on the domain cannot be found.
    """

    if domain is S.EmptySet:
        return S.EmptySet

    period = periodicity(f, symbol)
    if period == S.Zero:
        # the expression is constant wrt symbol
        return FiniteSet(f.expand())

    from sympy.series.limits import limit
    from sympy.solvers.solveset import solveset

    if period is not None:
        if isinstance(domain, Interval):
            if (domain.inf - domain.sup).is_infinite:
                domain = Interval(0, period)
        elif isinstance(domain, Union):
            for sub_dom in domain.args:
                if isinstance(sub_dom, Interval) and \
                ((sub_dom.inf - sub_dom.sup).is_infinite):
                    domain = Interval(0, period)

    intervals = continuous_domain(f, symbol, domain)
    range_int = S.EmptySet
    if isinstance(intervals,(Interval, FiniteSet)):
        interval_iter = (intervals,)
    elif isinstance(intervals, Union):
        interval_iter = intervals.args
    else:
        raise NotImplementedError("Unable to find range for the given domain.")

    for interval in interval_iter:
        if isinstance(interval, FiniteSet):
            for singleton in interval:
                if singleton in domain:
                    range_int += FiniteSet(f.subs(symbol, singleton))
        elif isinstance(interval, Interval):
            vals = S.EmptySet
            critical_values = S.EmptySet
            bounds = ((interval.left_open, interval.inf, '+'),
                   (interval.right_open, interval.sup, '-'))

            for is_open, limit_point, direction in bounds:
                if is_open:
                    critical_values += FiniteSet(limit(f, symbol, limit_point, direction))
                    vals += critical_values
                else:
                    vals += FiniteSet(f.subs(symbol, limit_point))

            critical_points = solveset(f.diff(symbol), symbol, interval)

            if not iterable(critical_points):
                raise NotImplementedError(
                        'Unable to find critical points for {}'.format(f))
            if isinstance(critical_points, ImageSet):
                raise NotImplementedError(
                        'Infinite number of critical points for {}'.format(f))

            for critical_point in critical_points:
                vals += FiniteSet(f.subs(symbol, critical_point))

            left_open, right_open = False, False

            if critical_values is not S.EmptySet:
                if critical_values.inf == vals.inf:
                    left_open = True

                if critical_values.sup == vals.sup:
                    right_open = True

            range_int += Interval(vals.inf, vals.sup, left_open, right_open)
        else:
            raise NotImplementedError("Unable to find range for the given domain.")

    return range_int


def not_empty_in(finset_intersection, *syms):
    """
    Finds the domain of the functions in ``finset_intersection`` in which the
    ``finite_set`` is not-empty.

    Parameters
    ==========

    finset_intersection : Intersection of FiniteSet
        The unevaluated intersection of FiniteSet containing
        real-valued functions with Union of Sets
    syms : Tuple of symbols
        Symbol for which domain is to be found

    Raises
    ======

    NotImplementedError
        The algorithms to find the non-emptiness of the given FiniteSet are
        not yet implemented.
    ValueError
        The input is not valid.
    RuntimeError
        It is a bug, please report it to the github issue tracker
        (https://github.com/sympy/sympy/issues).

    Examples
    ========

    >>> from sympy import FiniteSet, Interval, not_empty_in, oo
    >>> from sympy.abc import x
    >>> not_empty_in(FiniteSet(x/2).intersect(Interval(0, 1)), x)
    Interval(0, 2)
    >>> not_empty_in(FiniteSet(x, x**2).intersect(Interval(1, 2)), x)
    Union(Interval(1, 2), Interval(-sqrt(2), -1))
    >>> not_empty_in(FiniteSet(x**2/(x + 2)).intersect(Interval(1, oo)), x)
    Union(Interval.Lopen(-2, -1), Interval(2, oo))
    """

    # TODO: handle piecewise defined functions
    # TODO: handle transcendental functions
    # TODO: handle multivariate functions
    if len(syms) == 0:
        raise ValueError("One or more symbols must be given in syms.")

    if finset_intersection is S.EmptySet:
        return S.EmptySet

    if isinstance(finset_intersection, Union):
        elm_in_sets = finset_intersection.args[0]
        return Union(not_empty_in(finset_intersection.args[1], *syms),
                     elm_in_sets)

    if isinstance(finset_intersection, FiniteSet):
        finite_set = finset_intersection
        _sets = S.Reals
    else:
        finite_set = finset_intersection.args[1]
        _sets = finset_intersection.args[0]

    if not isinstance(finite_set, FiniteSet):
        raise ValueError('A FiniteSet must be given, not %s: %s' %
                         (type(finite_set), finite_set))

    if len(syms) == 1:
        symb = syms[0]
    else:
        raise NotImplementedError('more than one variables %s not handled' %
                                  (syms,))

    def elm_domain(expr, intrvl):
        """ Finds the domain of an expression in any given interval """
        from sympy.solvers.solveset import solveset

        _start = intrvl.start
        _end = intrvl.end
        _singularities = solveset(expr.as_numer_denom()[1], symb,
                                  domain=S.Reals)

        if intrvl.right_open:
            if _end is S.Infinity:
                _domain1 = S.Reals
            else:
                _domain1 = solveset(expr < _end, symb, domain=S.Reals)
        else:
            _domain1 = solveset(expr <= _end, symb, domain=S.Reals)

        if intrvl.left_open:
            if _start is S.NegativeInfinity:
                _domain2 = S.Reals
            else:
                _domain2 = solveset(expr > _start, symb, domain=S.Reals)
        else:
            _domain2 = solveset(expr >= _start, symb, domain=S.Reals)

        # domain in the interval
        expr_with_sing = Intersection(_domain1, _domain2)
        expr_domain = Complement(expr_with_sing, _singularities)
        return expr_domain

    if isinstance(_sets, Interval):
        return Union(*[elm_domain(element, _sets) for element in finite_set])

    if isinstance(_sets, Union):
        _domain = S.EmptySet
        for intrvl in _sets.args:
            _domain_element = Union(*[elm_domain(element, intrvl)
                                      for element in finite_set])
            _domain = Union(_domain, _domain_element)
        return _domain


def periodicity(f, symbol, check=False):
    """
    Tests the given function for periodicity in the given symbol.

    Parameters
    ==========

    f : :py:class:`~.Expr`
        The concerned function.
    symbol : :py:class:`~.Symbol`
        The variable for which the period is to be determined.
    check : bool, optional
        The flag to verify whether the value being returned is a period or not.

    Returns
    =======

    period
        The period of the function is returned.
        ``None`` is returned when the function is aperiodic or has a complex period.
        The value of $0$ is returned as the period of a constant function.

    Raises
    ======

    NotImplementedError
        The value of the period computed cannot be verified.


    Notes
    =====

    Currently, we do not support functions with a complex period.
    The period of functions having complex periodic values such
    as ``exp``, ``sinh`` is evaluated to ``None``.

    The value returned might not be the "fundamental" period of the given
    function i.e. it may not be the smallest periodic value of the function.

    The verification of the period through the ``check`` flag is not reliable
    due to internal simplification of the given expression. Hence, it is set
    to ``False`` by default.

    Examples
    ========
    >>> from sympy import periodicity, Symbol, sin, cos, tan, exp
    >>> x = Symbol('x')
    >>> f = sin(x) + sin(2*x) + sin(3*x)
    >>> periodicity(f, x)
    2*pi
    >>> periodicity(sin(x)*cos(x), x)
    pi
    >>> periodicity(exp(tan(2*x) - 1), x)
    pi/2
    >>> periodicity(sin(4*x)**cos(2*x), x)
    pi
    >>> periodicity(exp(x), x)
    """
    if symbol.kind is not NumberKind:
        raise NotImplementedError("Cannot use symbol of kind %s" % symbol.kind)
    temp = Dummy('x', real=True)
    f = f.subs(symbol, temp)
    symbol = temp

    def _check(orig_f, period):
        '''Return the checked period or raise an error.'''
        new_f = orig_f.subs(symbol, symbol + period)
        if new_f.equals(orig_f):
            return period
        else:
            raise NotImplementedError(filldedent('''
                The period of the given function cannot be verified.
                When `%s` was replaced with `%s + %s` in `%s`, the result
                was `%s` which was not recognized as being the same as
                the original function.
                So either the period was wrong or the two forms were
                not recognized as being equal.
                Set check=False to obtain the value.''' %
                (symbol, symbol, period, orig_f, new_f)))

    orig_f = f
    period = None

    if isinstance(f, Relational):
        f = f.lhs - f.rhs

    f = f.simplify()

    if symbol not in f.free_symbols:
        return S.Zero

    if isinstance(f, TrigonometricFunction):
        try:
            period = f.period(symbol)
        except NotImplementedError:
            pass

    if isinstance(f, Abs):
        arg = f.args[0]
        if isinstance(arg, (sec, csc, cos)):
            # all but tan and cot might have a
            # a period that is half as large
            # so recast as sin
            arg = sin(arg.args[0])
        period = periodicity(arg, symbol)
        if period is not None and isinstance(arg, sin):
            # the argument of Abs was a trigonometric other than
            # cot or tan; test to see if the half-period
            # is valid. Abs(arg) has behaviour equivalent to
            # orig_f, so use that for test:
            orig_f = Abs(arg)
            try:
                return _check(orig_f, period/2)
            except NotImplementedError as err:
                if check:
                    raise NotImplementedError(err)
            # else let new orig_f and period be
            # checked below

    if isinstance(f, exp) or (f.is_Pow and f.base == S.Exp1):
        f = Pow(S.Exp1, expand_mul(f.exp))
        if im(f) != 0:
            period_real = periodicity(re(f), symbol)
            period_imag = periodicity(im(f), symbol)
            if period_real is not None and period_imag is not None:
                period = lcim([period_real, period_imag])

    if f.is_Pow and f.base != S.Exp1:
        base, expo = f.args
        base_has_sym = base.has(symbol)
        expo_has_sym = expo.has(symbol)

        if base_has_sym and not expo_has_sym:
            period = periodicity(base, symbol)

        elif expo_has_sym and not base_has_sym:
            period = periodicity(expo, symbol)

        else:
            period = _periodicity(f.args, symbol)

    elif f.is_Mul:
        coeff, g = f.as_independent(symbol, as_Add=False)
        if isinstance(g, TrigonometricFunction) or not equal_valued(coeff, 1):
            period = periodicity(g, symbol)
        else:
            period = _periodicity(g.args, symbol)

    elif f.is_Add:
        k, g = f.as_independent(symbol)
        if k is not S.Zero:
            return periodicity(g, symbol)

        period = _periodicity(g.args, symbol)

    elif isinstance(f, Mod):
        a, n = f.args

        if a == symbol:
            period = n
        elif isinstance(a, TrigonometricFunction):
            period = periodicity(a, symbol)
        #check if 'f' is linear in 'symbol'
        elif (a.is_polynomial(symbol) and degree(a, symbol) == 1 and
            symbol not in n.free_symbols):
                period = Abs(n / a.diff(symbol))

    elif isinstance(f, Piecewise):
        pass  # not handling Piecewise yet as the return type is not favorable

    elif period is None:
        from sympy.solvers.decompogen import compogen, decompogen
        g_s = decompogen(f, symbol)
        num_of_gs = len(g_s)
        if num_of_gs > 1:
            for index, g in enumerate(reversed(g_s)):
                start_index = num_of_gs - 1 - index
                g = compogen(g_s[start_index:], symbol)
                if g not in (orig_f, f): # Fix for issue 12620
                    period = periodicity(g, symbol)
                    if period is not None:
                        break

    if period is not None:
        if check:
            return _check(orig_f, period)
        return period

    return None


def _periodicity(args, symbol):
    """
    Helper for `periodicity` to find the period of a list of simpler
    functions.
    It uses the `lcim` method to find the least common period of
    all the functions.

    Parameters
    ==========

    args : Tuple of :py:class:`~.Symbol`
        All the symbols present in a function.

    symbol : :py:class:`~.Symbol`
        The symbol over which the function is to be evaluated.

    Returns
    =======

    period
        The least common period of the function for all the symbols
        of the function.
        ``None`` if for at least one of the symbols the function is aperiodic.

    """
    periods = []
    for f in args:
        period = periodicity(f, symbol)
        if period is None:
            return None

        if period is not S.Zero:
            periods.append(period)

    if len(periods) > 1:
        return lcim(periods)

    if periods:
        return periods[0]


def lcim(numbers):
    """Returns the least common integral multiple of a list of numbers.

    The numbers can be rational or irrational or a mixture of both.
    `None` is returned for incommensurable numbers.

    Parameters
    ==========

    numbers : list
        Numbers (rational and/or irrational) for which lcim is to be found.

    Returns
    =======

    number
        lcim if it exists, otherwise ``None`` for incommensurable numbers.

    Examples
    ========

    >>> from sympy.calculus.util import lcim
    >>> from sympy import S, pi
    >>> lcim([S(1)/2, S(3)/4, S(5)/6])
    15/2
    >>> lcim([2*pi, 3*pi, pi, pi/2])
    6*pi
    >>> lcim([S(1), 2*pi])
    """
    result = None
    if all(num.is_irrational for num in numbers):
        factorized_nums = [num.factor() for num in numbers]
        factors_num = [num.as_coeff_Mul() for num in factorized_nums]
        term = factors_num[0][1]
        if all(factor == term for coeff, factor in factors_num):
            common_term = term
            coeffs = [coeff for coeff, factor in factors_num]
            result = lcm_list(coeffs) * common_term

    elif all(num.is_rational for num in numbers):
        result = lcm_list(numbers)

    else:
        pass

    return result

def is_convex(f, *syms, domain=S.Reals):
    r"""Determines the  convexity of the function passed in the argument.

    Parameters
    ==========

    f : :py:class:`~.Expr`
        The concerned function.
    syms : Tuple of :py:class:`~.Symbol`
        The variables with respect to which the convexity is to be determined.
    domain : :py:class:`~.Interval`, optional
        The domain over which the convexity of the function has to be checked.
        If unspecified, S.Reals will be the default domain.

    Returns
    =======

    bool
        The method returns ``True`` if the function is convex otherwise it
        returns ``False``.

    Raises
    ======

    NotImplementedError
        The check for the convexity of multivariate functions is not implemented yet.

    Notes
    =====

    To determine concavity of a function pass `-f` as the concerned function.
    To determine logarithmic convexity of a function pass `\log(f)` as
    concerned function.
    To determine logarithmic concavity of a function pass `-\log(f)` as
    concerned function.

    Currently, convexity check of multivariate functions is not handled.

    Examples
    ========

    >>> from sympy import is_convex, symbols, exp, oo, Interval
    >>> x = symbols('x')
    >>> is_convex(exp(x), x)
    True
    >>> is_convex(x**3, x, domain = Interval(-1, oo))
    False
    >>> is_convex(1/x**2, x, domain=Interval.open(0, oo))
    True

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Convex_function
    .. [2] http://www.ifp.illinois.edu/~angelia/L3_convfunc.pdf
    .. [3] https://en.wikipedia.org/wiki/Logarithmically_convex_function
    .. [4] https://en.wikipedia.org/wiki/Logarithmically_concave_function
    .. [5] https://en.wikipedia.org/wiki/Concave_function

    """
    if len(syms) > 1 :
        return hessian(f, syms).is_positive_semidefinite
    from sympy.solvers.inequalities import solve_univariate_inequality
    f = _sympify(f)
    var = syms[0]
    if any(s in domain for s in singularities(f, var)):
        return False
    condition = f.diff(var, 2) < 0
    if solve_univariate_inequality(condition, var, False, domain):
        return False
    return True


def stationary_points(f, symbol, domain=S.Reals):
    """
    Returns the stationary points of a function (where derivative of the
    function is 0) in the given domain.

    Parameters
    ==========

    f : :py:class:`~.Expr`
        The concerned function.
    symbol : :py:class:`~.Symbol`
        The variable for which the stationary points are to be determined.
    domain : :py:class:`~.Interval`
        The domain over which the stationary points have to be checked.
        If unspecified, ``S.Reals`` will be the default domain.

    Returns
    =======

    Set
        A set of stationary points for the function. If there are no
        stationary point, an :py:class:`~.EmptySet` is returned.

    Examples
    ========

    >>> from sympy import Interval, Symbol, S, sin, pi, pprint, stationary_points
    >>> x = Symbol('x')

    >>> stationary_points(1/x, x, S.Reals)
    EmptySet

    >>> pprint(stationary_points(sin(x), x), use_unicode=False)
              pi                              3*pi
    {2*n*pi + -- | n in Integers} U {2*n*pi + ---- | n in Integers}
              2                                2

    >>> stationary_points(sin(x),x, Interval(0, 4*pi))
    {pi/2, 3*pi/2, 5*pi/2, 7*pi/2}

    """
    from sympy.solvers.solveset import solveset

    if domain is S.EmptySet:
        return S.EmptySet

    domain = continuous_domain(f, symbol, domain)
    set = solveset(diff(f, symbol), symbol, domain)

    return set


def maximum(f, symbol, domain=S.Reals):
    """
    Returns the maximum value of a function in the given domain.

    Parameters
    ==========

    f : :py:class:`~.Expr`
        The concerned function.
    symbol : :py:class:`~.Symbol`
        The variable for maximum value needs to be determined.
    domain : :py:class:`~.Interval`
        The domain over which the maximum have to be checked.
        If unspecified, then the global maximum is returned.

    Returns
    =======

    number
        Maximum value of the function in given domain.

    Examples
    ========

    >>> from sympy import Interval, Symbol, S, sin, cos, pi, maximum
    >>> x = Symbol('x')

    >>> f = -x**2 + 2*x + 5
    >>> maximum(f, x, S.Reals)
    6

    >>> maximum(sin(x), x, Interval(-pi, pi/4))
    sqrt(2)/2

    >>> maximum(sin(x)*cos(x), x)
    1/2

    """
    if isinstance(symbol, Symbol):
        if domain is S.EmptySet:
            raise ValueError("Maximum value not defined for empty domain.")

        return function_range(f, symbol, domain).sup
    else:
        raise ValueError("%s is not a valid symbol." % symbol)


def minimum(f, symbol, domain=S.Reals):
    """
    Returns the minimum value of a function in the given domain.

    Parameters
    ==========

    f : :py:class:`~.Expr`
        The concerned function.
    symbol : :py:class:`~.Symbol`
        The variable for minimum value needs to be determined.
    domain : :py:class:`~.Interval`
        The domain over which the minimum have to be checked.
        If unspecified, then the global minimum is returned.

    Returns
    =======

    number
        Minimum value of the function in the given domain.

    Examples
    ========

    >>> from sympy import Interval, Symbol, S, sin, cos, minimum
    >>> x = Symbol('x')

    >>> f = x**2 + 2*x + 5
    >>> minimum(f, x, S.Reals)
    4

    >>> minimum(sin(x), x, Interval(2, 3))
    sin(3)

    >>> minimum(sin(x)*cos(x), x)
    -1/2

    """
    if isinstance(symbol, Symbol):
        if domain is S.EmptySet:
            raise ValueError("Minimum value not defined for empty domain.")

        return function_range(f, symbol, domain).inf
    else:
        raise ValueError("%s is not a valid symbol." % symbol)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\reshape\concat.py ===
"""
Concat routines.
"""
from __future__ import annotations

from collections import abc
from typing import (
    TYPE_CHECKING,
    Callable,
    Literal,
    cast,
    overload,
)
import warnings

import numpy as np

from pandas._config import using_copy_on_write

from pandas.util._decorators import cache_readonly
from pandas.util._exceptions import find_stack_level

from pandas.core.dtypes.common import (
    is_bool,
    is_iterator,
)
from pandas.core.dtypes.concat import concat_compat
from pandas.core.dtypes.generic import (
    ABCDataFrame,
    ABCSeries,
)
from pandas.core.dtypes.missing import isna

from pandas.core.arrays.categorical import (
    factorize_from_iterable,
    factorize_from_iterables,
)
import pandas.core.common as com
from pandas.core.indexes.api import (
    Index,
    MultiIndex,
    all_indexes_same,
    default_index,
    ensure_index,
    get_objs_combined_axis,
    get_unanimous_names,
)
from pandas.core.internals import concatenate_managers

if TYPE_CHECKING:
    from collections.abc import (
        Hashable,
        Iterable,
        Mapping,
    )

    from pandas._typing import (
        Axis,
        AxisInt,
        HashableT,
    )

    from pandas import (
        DataFrame,
        Series,
    )

# ---------------------------------------------------------------------
# Concatenate DataFrame objects


@overload
def concat(
    objs: Iterable[DataFrame] | Mapping[HashableT, DataFrame],
    *,
    axis: Literal[0, "index"] = ...,
    join: str = ...,
    ignore_index: bool = ...,
    keys: Iterable[Hashable] | None = ...,
    levels=...,
    names: list[HashableT] | None = ...,
    verify_integrity: bool = ...,
    sort: bool = ...,
    copy: bool | None = ...,
) -> DataFrame:
    ...


@overload
def concat(
    objs: Iterable[Series] | Mapping[HashableT, Series],
    *,
    axis: Literal[0, "index"] = ...,
    join: str = ...,
    ignore_index: bool = ...,
    keys: Iterable[Hashable] | None = ...,
    levels=...,
    names: list[HashableT] | None = ...,
    verify_integrity: bool = ...,
    sort: bool = ...,
    copy: bool | None = ...,
) -> Series:
    ...


@overload
def concat(
    objs: Iterable[Series | DataFrame] | Mapping[HashableT, Series | DataFrame],
    *,
    axis: Literal[0, "index"] = ...,
    join: str = ...,
    ignore_index: bool = ...,
    keys: Iterable[Hashable] | None = ...,
    levels=...,
    names: list[HashableT] | None = ...,
    verify_integrity: bool = ...,
    sort: bool = ...,
    copy: bool | None = ...,
) -> DataFrame | Series:
    ...


@overload
def concat(
    objs: Iterable[Series | DataFrame] | Mapping[HashableT, Series | DataFrame],
    *,
    axis: Literal[1, "columns"],
    join: str = ...,
    ignore_index: bool = ...,
    keys: Iterable[Hashable] | None = ...,
    levels=...,
    names: list[HashableT] | None = ...,
    verify_integrity: bool = ...,
    sort: bool = ...,
    copy: bool | None = ...,
) -> DataFrame:
    ...


@overload
def concat(
    objs: Iterable[Series | DataFrame] | Mapping[HashableT, Series | DataFrame],
    *,
    axis: Axis = ...,
    join: str = ...,
    ignore_index: bool = ...,
    keys: Iterable[Hashable] | None = ...,
    levels=...,
    names: list[HashableT] | None = ...,
    verify_integrity: bool = ...,
    sort: bool = ...,
    copy: bool | None = ...,
) -> DataFrame | Series:
    ...


def concat(
    objs: Iterable[Series | DataFrame] | Mapping[HashableT, Series | DataFrame],
    *,
    axis: Axis = 0,
    join: str = "outer",
    ignore_index: bool = False,
    keys: Iterable[Hashable] | None = None,
    levels=None,
    names: list[HashableT] | None = None,
    verify_integrity: bool = False,
    sort: bool = False,
    copy: bool | None = None,
) -> DataFrame | Series:
    """
    Concatenate pandas objects along a particular axis.

    Allows optional set logic along the other axes.

    Can also add a layer of hierarchical indexing on the concatenation axis,
    which may be useful if the labels are the same (or overlapping) on
    the passed axis number.

    Parameters
    ----------
    objs : a sequence or mapping of Series or DataFrame objects
        If a mapping is passed, the sorted keys will be used as the `keys`
        argument, unless it is passed, in which case the values will be
        selected (see below). Any None objects will be dropped silently unless
        they are all None in which case a ValueError will be raised.
    axis : {0/'index', 1/'columns'}, default 0
        The axis to concatenate along.
    join : {'inner', 'outer'}, default 'outer'
        How to handle indexes on other axis (or axes).
    ignore_index : bool, default False
        If True, do not use the index values along the concatenation axis. The
        resulting axis will be labeled 0, ..., n - 1. This is useful if you are
        concatenating objects where the concatenation axis does not have
        meaningful indexing information. Note the index values on the other
        axes are still respected in the join.
    keys : sequence, default None
        If multiple levels passed, should contain tuples. Construct
        hierarchical index using the passed keys as the outermost level.
    levels : list of sequences, default None
        Specific levels (unique values) to use for constructing a
        MultiIndex. Otherwise they will be inferred from the keys.
    names : list, default None
        Names for the levels in the resulting hierarchical index.
    verify_integrity : bool, default False
        Check whether the new concatenated axis contains duplicates. This can
        be very expensive relative to the actual data concatenation.
    sort : bool, default False
        Sort non-concatenation axis if it is not already aligned. One exception to
        this is when the non-concatentation axis is a DatetimeIndex and join='outer'
        and the axis is not already aligned. In that case, the non-concatenation
        axis is always sorted lexicographically.
    copy : bool, default True
        If False, do not copy data unnecessarily.

    Returns
    -------
    object, type of objs
        When concatenating all ``Series`` along the index (axis=0), a
        ``Series`` is returned. When ``objs`` contains at least one
        ``DataFrame``, a ``DataFrame`` is returned. When concatenating along
        the columns (axis=1), a ``DataFrame`` is returned.

    See Also
    --------
    DataFrame.join : Join DataFrames using indexes.
    DataFrame.merge : Merge DataFrames by indexes or columns.

    Notes
    -----
    The keys, levels, and names arguments are all optional.

    A walkthrough of how this method fits in with other tools for combining
    pandas objects can be found `here
    <https://pandas.pydata.org/pandas-docs/stable/user_guide/merging.html>`__.

    It is not recommended to build DataFrames by adding single rows in a
    for loop. Build a list of rows and make a DataFrame in a single concat.

    Examples
    --------
    Combine two ``Series``.

    >>> s1 = pd.Series(['a', 'b'])
    >>> s2 = pd.Series(['c', 'd'])
    >>> pd.concat([s1, s2])
    0    a
    1    b
    0    c
    1    d
    dtype: object

    Clear the existing index and reset it in the result
    by setting the ``ignore_index`` option to ``True``.

    >>> pd.concat([s1, s2], ignore_index=True)
    0    a
    1    b
    2    c
    3    d
    dtype: object

    Add a hierarchical index at the outermost level of
    the data with the ``keys`` option.

    >>> pd.concat([s1, s2], keys=['s1', 's2'])
    s1  0    a
        1    b
    s2  0    c
        1    d
    dtype: object

    Label the index keys you create with the ``names`` option.

    >>> pd.concat([s1, s2], keys=['s1', 's2'],
    ...           names=['Series name', 'Row ID'])
    Series name  Row ID
    s1           0         a
                 1         b
    s2           0         c
                 1         d
    dtype: object

    Combine two ``DataFrame`` objects with identical columns.

    >>> df1 = pd.DataFrame([['a', 1], ['b', 2]],
    ...                    columns=['letter', 'number'])
    >>> df1
      letter  number
    0      a       1
    1      b       2
    >>> df2 = pd.DataFrame([['c', 3], ['d', 4]],
    ...                    columns=['letter', 'number'])
    >>> df2
      letter  number
    0      c       3
    1      d       4
    >>> pd.concat([df1, df2])
      letter  number
    0      a       1
    1      b       2
    0      c       3
    1      d       4

    Combine ``DataFrame`` objects with overlapping columns
    and return everything. Columns outside the intersection will
    be filled with ``NaN`` values.

    >>> df3 = pd.DataFrame([['c', 3, 'cat'], ['d', 4, 'dog']],
    ...                    columns=['letter', 'number', 'animal'])
    >>> df3
      letter  number animal
    0      c       3    cat
    1      d       4    dog
    >>> pd.concat([df1, df3], sort=False)
      letter  number animal
    0      a       1    NaN
    1      b       2    NaN
    0      c       3    cat
    1      d       4    dog

    Combine ``DataFrame`` objects with overlapping columns
    and return only those that are shared by passing ``inner`` to
    the ``join`` keyword argument.

    >>> pd.concat([df1, df3], join="inner")
      letter  number
    0      a       1
    1      b       2
    0      c       3
    1      d       4

    Combine ``DataFrame`` objects horizontally along the x axis by
    passing in ``axis=1``.

    >>> df4 = pd.DataFrame([['bird', 'polly'], ['monkey', 'george']],
    ...                    columns=['animal', 'name'])
    >>> pd.concat([df1, df4], axis=1)
      letter  number  animal    name
    0      a       1    bird   polly
    1      b       2  monkey  george

    Prevent the result from including duplicate index values with the
    ``verify_integrity`` option.

    >>> df5 = pd.DataFrame([1], index=['a'])
    >>> df5
       0
    a  1
    >>> df6 = pd.DataFrame([2], index=['a'])
    >>> df6
       0
    a  2
    >>> pd.concat([df5, df6], verify_integrity=True)
    Traceback (most recent call last):
        ...
    ValueError: Indexes have overlapping values: ['a']

    Append a single row to the end of a ``DataFrame`` object.

    >>> df7 = pd.DataFrame({'a': 1, 'b': 2}, index=[0])
    >>> df7
        a   b
    0   1   2
    >>> new_row = pd.Series({'a': 3, 'b': 4})
    >>> new_row
    a    3
    b    4
    dtype: int64
    >>> pd.concat([df7, new_row.to_frame().T], ignore_index=True)
        a   b
    0   1   2
    1   3   4
    """
    if copy is None:
        if using_copy_on_write():
            copy = False
        else:
            copy = True
    elif copy and using_copy_on_write():
        copy = False

    op = _Concatenator(
        objs,
        axis=axis,
        ignore_index=ignore_index,
        join=join,
        keys=keys,
        levels=levels,
        names=names,
        verify_integrity=verify_integrity,
        copy=copy,
        sort=sort,
    )

    return op.get_result()


class _Concatenator:
    """
    Orchestrates a concatenation operation for BlockManagers
    """

    sort: bool

    def __init__(
        self,
        objs: Iterable[Series | DataFrame] | Mapping[HashableT, Series | DataFrame],
        axis: Axis = 0,
        join: str = "outer",
        keys: Iterable[Hashable] | None = None,
        levels=None,
        names: list[HashableT] | None = None,
        ignore_index: bool = False,
        verify_integrity: bool = False,
        copy: bool = True,
        sort: bool = False,
    ) -> None:
        if isinstance(objs, (ABCSeries, ABCDataFrame, str)):
            raise TypeError(
                "first argument must be an iterable of pandas "
                f'objects, you passed an object of type "{type(objs).__name__}"'
            )

        if join == "outer":
            self.intersect = False
        elif join == "inner":
            self.intersect = True
        else:  # pragma: no cover
            raise ValueError(
                "Only can inner (intersect) or outer (union) join the other axis"
            )

        if not is_bool(sort):
            raise ValueError(
                f"The 'sort' keyword only accepts boolean values; {sort} was passed."
            )
        # Incompatible types in assignment (expression has type "Union[bool, bool_]",
        # variable has type "bool")
        self.sort = sort  # type: ignore[assignment]

        self.ignore_index = ignore_index
        self.verify_integrity = verify_integrity
        self.copy = copy

        objs, keys = self._clean_keys_and_objs(objs, keys)

        # figure out what our result ndim is going to be
        ndims = self._get_ndims(objs)
        sample, objs = self._get_sample_object(objs, ndims, keys, names, levels)

        # Standardize axis parameter to int
        if sample.ndim == 1:
            from pandas import DataFrame

            axis = DataFrame._get_axis_number(axis)
            self._is_frame = False
            self._is_series = True
        else:
            axis = sample._get_axis_number(axis)
            self._is_frame = True
            self._is_series = False

            # Need to flip BlockManager axis in the DataFrame special case
            axis = sample._get_block_manager_axis(axis)

        # if we have mixed ndims, then convert to highest ndim
        # creating column numbers as needed
        if len(ndims) > 1:
            objs = self._sanitize_mixed_ndim(objs, sample, ignore_index, axis)

        self.objs = objs

        # note: this is the BlockManager axis (since DataFrame is transposed)
        self.bm_axis = axis
        self.axis = 1 - self.bm_axis if self._is_frame else 0
        self.keys = keys
        self.names = names or getattr(keys, "names", None)
        self.levels = levels

    def _get_ndims(self, objs: list[Series | DataFrame]) -> set[int]:
        # figure out what our result ndim is going to be
        ndims = set()
        for obj in objs:
            if not isinstance(obj, (ABCSeries, ABCDataFrame)):
                msg = (
                    f"cannot concatenate object of type '{type(obj)}'; "
                    "only Series and DataFrame objs are valid"
                )
                raise TypeError(msg)

            ndims.add(obj.ndim)
        return ndims

    def _clean_keys_and_objs(
        self,
        objs: Iterable[Series | DataFrame] | Mapping[HashableT, Series | DataFrame],
        keys,
    ) -> tuple[list[Series | DataFrame], Index | None]:
        if isinstance(objs, abc.Mapping):
            if keys is None:
                keys = list(objs.keys())
            objs_list = [objs[k] for k in keys]
        else:
            objs_list = list(objs)

        if len(objs_list) == 0:
            raise ValueError("No objects to concatenate")

        if keys is None:
            objs_list = list(com.not_none(*objs_list))
        else:
            # GH#1649
            clean_keys = []
            clean_objs = []
            if is_iterator(keys):
                keys = list(keys)
            if len(keys) != len(objs_list):
                # GH#43485
                warnings.warn(
                    "The behavior of pd.concat with len(keys) != len(objs) is "
                    "deprecated. In a future version this will raise instead of "
                    "truncating to the smaller of the two sequences",
                    FutureWarning,
                    stacklevel=find_stack_level(),
                )
            for k, v in zip(keys, objs_list):
                if v is None:
                    continue
                clean_keys.append(k)
                clean_objs.append(v)
            objs_list = clean_objs

            if isinstance(keys, MultiIndex):
                # TODO: retain levels?
                keys = type(keys).from_tuples(clean_keys, names=keys.names)
            else:
                name = getattr(keys, "name", None)
                keys = Index(clean_keys, name=name, dtype=getattr(keys, "dtype", None))

        if len(objs_list) == 0:
            raise ValueError("All objects passed were None")

        return objs_list, keys

    def _get_sample_object(
        self,
        objs: list[Series | DataFrame],
        ndims: set[int],
        keys,
        names,
        levels,
    ) -> tuple[Series | DataFrame, list[Series | DataFrame]]:
        # get the sample
        # want the highest ndim that we have, and must be non-empty
        # unless all objs are empty
        sample: Series | DataFrame | None = None
        if len(ndims) > 1:
            max_ndim = max(ndims)
            for obj in objs:
                if obj.ndim == max_ndim and np.sum(obj.shape):
                    sample = obj
                    break

        else:
            # filter out the empties if we have not multi-index possibilities
            # note to keep empty Series as it affect to result columns / name
            non_empties = [obj for obj in objs if sum(obj.shape) > 0 or obj.ndim == 1]

            if len(non_empties) and (
                keys is None and names is None and levels is None and not self.intersect
            ):
                objs = non_empties
                sample = objs[0]

        if sample is None:
            sample = objs[0]
        return sample, objs

    def _sanitize_mixed_ndim(
        self,
        objs: list[Series | DataFrame],
        sample: Series | DataFrame,
        ignore_index: bool,
        axis: AxisInt,
    ) -> list[Series | DataFrame]:
        # if we have mixed ndims, then convert to highest ndim
        # creating column numbers as needed

        new_objs = []

        current_column = 0
        max_ndim = sample.ndim
        for obj in objs:
            ndim = obj.ndim
            if ndim == max_ndim:
                pass

            elif ndim != max_ndim - 1:
                raise ValueError(
                    "cannot concatenate unaligned mixed dimensional NDFrame objects"
                )

            else:
                name = getattr(obj, "name", None)
                if ignore_index or name is None:
                    if axis == 1:
                        # doing a row-wise concatenation so need everything
                        # to line up
                        name = 0
                    else:
                        # doing a column-wise concatenation so need series
                        # to have unique names
                        name = current_column
                        current_column += 1

                obj = sample._constructor({name: obj}, copy=False)

            new_objs.append(obj)

        return new_objs

    def get_result(self):
        cons: Callable[..., DataFrame | Series]
        sample: DataFrame | Series

        # series only
        if self._is_series:
            sample = cast("Series", self.objs[0])

            # stack blocks
            if self.bm_axis == 0:
                name = com.consensus_name_attr(self.objs)
                cons = sample._constructor

                arrs = [ser._values for ser in self.objs]

                res = concat_compat(arrs, axis=0)

                new_index: Index
                if self.ignore_index:
                    # We can avoid surprisingly-expensive _get_concat_axis
                    new_index = default_index(len(res))
                else:
                    new_index = self.new_axes[0]

                mgr = type(sample._mgr).from_array(res, index=new_index)

                result = sample._constructor_from_mgr(mgr, axes=mgr.axes)
                result._name = name
                return result.__finalize__(self, method="concat")

            # combine as columns in a frame
            else:
                data = dict(zip(range(len(self.objs)), self.objs))

                # GH28330 Preserves subclassed objects through concat
                cons = sample._constructor_expanddim

                index, columns = self.new_axes
                df = cons(data, index=index, copy=self.copy)
                df.columns = columns
                return df.__finalize__(self, method="concat")

        # combine block managers
        else:
            sample = cast("DataFrame", self.objs[0])

            mgrs_indexers = []
            for obj in self.objs:
                indexers = {}
                for ax, new_labels in enumerate(self.new_axes):
                    # ::-1 to convert BlockManager ax to DataFrame ax
                    if ax == self.bm_axis:
                        # Suppress reindexing on concat axis
                        continue

                    # 1-ax to convert BlockManager axis to DataFrame axis
                    obj_labels = obj.axes[1 - ax]
                    if not new_labels.equals(obj_labels):
                        indexers[ax] = obj_labels.get_indexer(new_labels)

                mgrs_indexers.append((obj._mgr, indexers))

            new_data = concatenate_managers(
                mgrs_indexers, self.new_axes, concat_axis=self.bm_axis, copy=self.copy
            )
            if not self.copy and not using_copy_on_write():
                new_data._consolidate_inplace()

            out = sample._constructor_from_mgr(new_data, axes=new_data.axes)
            return out.__finalize__(self, method="concat")

    def _get_result_dim(self) -> int:
        if self._is_series and self.bm_axis == 1:
            return 2
        else:
            return self.objs[0].ndim

    @cache_readonly
    def new_axes(self) -> list[Index]:
        ndim = self._get_result_dim()
        return [
            self._get_concat_axis if i == self.bm_axis else self._get_comb_axis(i)
            for i in range(ndim)
        ]

    def _get_comb_axis(self, i: AxisInt) -> Index:
        data_axis = self.objs[0]._get_block_manager_axis(i)
        return get_objs_combined_axis(
            self.objs,
            axis=data_axis,
            intersect=self.intersect,
            sort=self.sort,
            copy=self.copy,
        )

    @cache_readonly
    def _get_concat_axis(self) -> Index:
        """
        Return index to be used along concatenation axis.
        """
        if self._is_series:
            if self.bm_axis == 0:
                indexes = [x.index for x in self.objs]
            elif self.ignore_index:
                idx = default_index(len(self.objs))
                return idx
            elif self.keys is None:
                names: list[Hashable] = [None] * len(self.objs)
                num = 0
                has_names = False
                for i, x in enumerate(self.objs):
                    if x.ndim != 1:
                        raise TypeError(
                            f"Cannot concatenate type 'Series' with "
                            f"object of type '{type(x).__name__}'"
                        )
                    if x.name is not None:
                        names[i] = x.name
                        has_names = True
                    else:
                        names[i] = num
                        num += 1
                if has_names:
                    return Index(names)
                else:
                    return default_index(len(self.objs))
            else:
                return ensure_index(self.keys).set_names(self.names)
        else:
            indexes = [x.axes[self.axis] for x in self.objs]

        if self.ignore_index:
            idx = default_index(sum(len(i) for i in indexes))
            return idx

        if self.keys is None:
            if self.levels is not None:
                raise ValueError("levels supported only when keys is not None")
            concat_axis = _concat_indexes(indexes)
        else:
            concat_axis = _make_concat_multiindex(
                indexes, self.keys, self.levels, self.names
            )

        self._maybe_check_integrity(concat_axis)

        return concat_axis

    def _maybe_check_integrity(self, concat_index: Index):
        if self.verify_integrity:
            if not concat_index.is_unique:
                overlap = concat_index[concat_index.duplicated()].unique()
                raise ValueError(f"Indexes have overlapping values: {overlap}")


def _concat_indexes(indexes) -> Index:
    return indexes[0].append(indexes[1:])


def _make_concat_multiindex(indexes, keys, levels=None, names=None) -> MultiIndex:
    if (levels is None and isinstance(keys[0], tuple)) or (
        levels is not None and len(levels) > 1
    ):
        zipped = list(zip(*keys))
        if names is None:
            names = [None] * len(zipped)

        if levels is None:
            _, levels = factorize_from_iterables(zipped)
        else:
            levels = [ensure_index(x) for x in levels]
    else:
        zipped = [keys]
        if names is None:
            names = [None]

        if levels is None:
            levels = [ensure_index(keys).unique()]
        else:
            levels = [ensure_index(x) for x in levels]

    for level in levels:
        if not level.is_unique:
            raise ValueError(f"Level values not unique: {level.tolist()}")

    if not all_indexes_same(indexes) or not all(level.is_unique for level in levels):
        codes_list = []

        # things are potentially different sizes, so compute the exact codes
        # for each level and pass those to MultiIndex.from_arrays

        for hlevel, level in zip(zipped, levels):
            to_concat = []
            if isinstance(hlevel, Index) and hlevel.equals(level):
                lens = [len(idx) for idx in indexes]
                codes_list.append(np.repeat(np.arange(len(hlevel)), lens))
            else:
                for key, index in zip(hlevel, indexes):
                    # Find matching codes, include matching nan values as equal.
                    mask = (isna(level) & isna(key)) | (level == key)
                    if not mask.any():
                        raise ValueError(f"Key {key} not in level {level}")
                    i = np.nonzero(mask)[0][0]

                    to_concat.append(np.repeat(i, len(index)))
                codes_list.append(np.concatenate(to_concat))

        concat_index = _concat_indexes(indexes)

        # these go at the end
        if isinstance(concat_index, MultiIndex):
            levels.extend(concat_index.levels)
            codes_list.extend(concat_index.codes)
        else:
            codes, categories = factorize_from_iterable(concat_index)
            levels.append(categories)
            codes_list.append(codes)

        if len(names) == len(levels):
            names = list(names)
        else:
            # make sure that all of the passed indices have the same nlevels
            if not len({idx.nlevels for idx in indexes}) == 1:
                raise AssertionError(
                    "Cannot concat indices that do not have the same number of levels"
                )

            # also copies
            names = list(names) + list(get_unanimous_names(*indexes))

        return MultiIndex(
            levels=levels, codes=codes_list, names=names, verify_integrity=False
        )

    new_index = indexes[0]
    n = len(new_index)
    kpieces = len(indexes)

    # also copies
    new_names = list(names)
    new_levels = list(levels)

    # construct codes
    new_codes = []

    # do something a bit more speedy

    for hlevel, level in zip(zipped, levels):
        hlevel_index = ensure_index(hlevel)
        mapped = level.get_indexer(hlevel_index)

        mask = mapped == -1
        if mask.any():
            raise ValueError(
                f"Values not found in passed level: {hlevel_index[mask]!s}"
            )

        new_codes.append(np.repeat(mapped, n))

    if isinstance(new_index, MultiIndex):
        new_levels.extend(new_index.levels)
        new_codes.extend([np.tile(lab, kpieces) for lab in new_index.codes])
    else:
        new_levels.append(new_index.unique())
        single_codes = new_index.unique().get_indexer(new_index)
        new_codes.append(np.tile(single_codes, kpieces))

    if len(new_names) < len(new_levels):
        new_names.extend(new_index.names)

    return MultiIndex(
        levels=new_levels, codes=new_codes, names=new_names, verify_integrity=False
    )