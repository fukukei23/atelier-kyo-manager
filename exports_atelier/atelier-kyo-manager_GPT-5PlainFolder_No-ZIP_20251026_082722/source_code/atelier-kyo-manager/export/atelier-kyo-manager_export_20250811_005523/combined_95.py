
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\orm\base.py ===
# orm/base.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

"""Constants and rudimental functions used throughout the ORM.

"""

from __future__ import annotations

from enum import Enum
import operator
import typing
from typing import Any
from typing import Callable
from typing import Dict
from typing import Generic
from typing import no_type_check
from typing import Optional
from typing import overload
from typing import Tuple
from typing import Type
from typing import TYPE_CHECKING
from typing import TypeVar
from typing import Union

from . import exc
from ._typing import insp_is_mapper
from .. import exc as sa_exc
from .. import inspection
from .. import util
from ..sql import roles
from ..sql.elements import SQLColumnExpression
from ..sql.elements import SQLCoreOperations
from ..util import FastIntFlag
from ..util.langhelpers import TypingOnly
from ..util.typing import Literal

if typing.TYPE_CHECKING:
    from ._typing import _EntityType
    from ._typing import _ExternalEntityType
    from ._typing import _InternalEntityType
    from .attributes import InstrumentedAttribute
    from .dynamic import AppenderQuery
    from .instrumentation import ClassManager
    from .interfaces import PropComparator
    from .mapper import Mapper
    from .state import InstanceState
    from .util import AliasedClass
    from .writeonly import WriteOnlyCollection
    from ..sql._typing import _ColumnExpressionArgument
    from ..sql._typing import _InfoType
    from ..sql.elements import ColumnElement
    from ..sql.operators import OperatorType

_T = TypeVar("_T", bound=Any)
_T_co = TypeVar("_T_co", bound=Any, covariant=True)

_O = TypeVar("_O", bound=object)


class LoaderCallableStatus(Enum):
    PASSIVE_NO_RESULT = 0
    """Symbol returned by a loader callable or other attribute/history
    retrieval operation when a value could not be determined, based
    on loader callable flags.
    """

    PASSIVE_CLASS_MISMATCH = 1
    """Symbol indicating that an object is locally present for a given
    primary key identity but it is not of the requested class.  The
    return value is therefore None and no SQL should be emitted."""

    ATTR_WAS_SET = 2
    """Symbol returned by a loader callable to indicate the
    retrieved value, or values, were assigned to their attributes
    on the target object.
    """

    ATTR_EMPTY = 3
    """Symbol used internally to indicate an attribute had no callable."""

    NO_VALUE = 4
    """Symbol which may be placed as the 'previous' value of an attribute,
    indicating no value was loaded for an attribute when it was modified,
    and flags indicated we were not to load it.
    """

    NEVER_SET = NO_VALUE
    """
    Synonymous with NO_VALUE

    .. versionchanged:: 1.4   NEVER_SET was merged with NO_VALUE

    """


(
    PASSIVE_NO_RESULT,
    PASSIVE_CLASS_MISMATCH,
    ATTR_WAS_SET,
    ATTR_EMPTY,
    NO_VALUE,
) = tuple(LoaderCallableStatus)

NEVER_SET = NO_VALUE


class PassiveFlag(FastIntFlag):
    """Bitflag interface that passes options onto loader callables"""

    NO_CHANGE = 0
    """No callables or SQL should be emitted on attribute access
    and no state should change
    """

    CALLABLES_OK = 1
    """Loader callables can be fired off if a value
    is not present.
    """

    SQL_OK = 2
    """Loader callables can emit SQL at least on scalar value attributes."""

    RELATED_OBJECT_OK = 4
    """Callables can use SQL to load related objects as well
    as scalar value attributes.
    """

    INIT_OK = 8
    """Attributes should be initialized with a blank
    value (None or an empty collection) upon get, if no other
    value can be obtained.
    """

    NON_PERSISTENT_OK = 16
    """Callables can be emitted if the parent is not persistent."""

    LOAD_AGAINST_COMMITTED = 32
    """Callables should use committed values as primary/foreign keys during a
    load.
    """

    NO_AUTOFLUSH = 64
    """Loader callables should disable autoflush."""

    NO_RAISE = 128
    """Loader callables should not raise any assertions"""

    DEFERRED_HISTORY_LOAD = 256
    """indicates special load of the previous value of an attribute"""

    INCLUDE_PENDING_MUTATIONS = 512

    # pre-packaged sets of flags used as inputs
    PASSIVE_OFF = (
        RELATED_OBJECT_OK | NON_PERSISTENT_OK | INIT_OK | CALLABLES_OK | SQL_OK
    )
    "Callables can be emitted in all cases."

    PASSIVE_RETURN_NO_VALUE = PASSIVE_OFF ^ INIT_OK
    """PASSIVE_OFF ^ INIT_OK"""

    PASSIVE_NO_INITIALIZE = PASSIVE_RETURN_NO_VALUE ^ CALLABLES_OK
    "PASSIVE_RETURN_NO_VALUE ^ CALLABLES_OK"

    PASSIVE_NO_FETCH = PASSIVE_OFF ^ SQL_OK
    "PASSIVE_OFF ^ SQL_OK"

    PASSIVE_NO_FETCH_RELATED = PASSIVE_OFF ^ RELATED_OBJECT_OK
    "PASSIVE_OFF ^ RELATED_OBJECT_OK"

    PASSIVE_ONLY_PERSISTENT = PASSIVE_OFF ^ NON_PERSISTENT_OK
    "PASSIVE_OFF ^ NON_PERSISTENT_OK"

    PASSIVE_MERGE = PASSIVE_OFF | NO_RAISE
    """PASSIVE_OFF | NO_RAISE

    Symbol used specifically for session.merge() and similar cases

    """


(
    NO_CHANGE,
    CALLABLES_OK,
    SQL_OK,
    RELATED_OBJECT_OK,
    INIT_OK,
    NON_PERSISTENT_OK,
    LOAD_AGAINST_COMMITTED,
    NO_AUTOFLUSH,
    NO_RAISE,
    DEFERRED_HISTORY_LOAD,
    INCLUDE_PENDING_MUTATIONS,
    PASSIVE_OFF,
    PASSIVE_RETURN_NO_VALUE,
    PASSIVE_NO_INITIALIZE,
    PASSIVE_NO_FETCH,
    PASSIVE_NO_FETCH_RELATED,
    PASSIVE_ONLY_PERSISTENT,
    PASSIVE_MERGE,
) = PassiveFlag.__members__.values()

DEFAULT_MANAGER_ATTR = "_sa_class_manager"
DEFAULT_STATE_ATTR = "_sa_instance_state"


class EventConstants(Enum):
    EXT_CONTINUE = 1
    EXT_STOP = 2
    EXT_SKIP = 3
    NO_KEY = 4
    """indicates an :class:`.AttributeEvent` event that did not have any
    key argument.

    .. versionadded:: 2.0

    """


EXT_CONTINUE, EXT_STOP, EXT_SKIP, NO_KEY = tuple(EventConstants)


class RelationshipDirection(Enum):
    """enumeration which indicates the 'direction' of a
    :class:`_orm.RelationshipProperty`.

    :class:`.RelationshipDirection` is accessible from the
    :attr:`_orm.Relationship.direction` attribute of
    :class:`_orm.RelationshipProperty`.

    """

    ONETOMANY = 1
    """Indicates the one-to-many direction for a :func:`_orm.relationship`.

    This symbol is typically used by the internals but may be exposed within
    certain API features.

    """

    MANYTOONE = 2
    """Indicates the many-to-one direction for a :func:`_orm.relationship`.

    This symbol is typically used by the internals but may be exposed within
    certain API features.

    """

    MANYTOMANY = 3
    """Indicates the many-to-many direction for a :func:`_orm.relationship`.

    This symbol is typically used by the internals but may be exposed within
    certain API features.

    """


ONETOMANY, MANYTOONE, MANYTOMANY = tuple(RelationshipDirection)


class InspectionAttrExtensionType(Enum):
    """Symbols indicating the type of extension that a
    :class:`.InspectionAttr` is part of."""


class NotExtension(InspectionAttrExtensionType):
    NOT_EXTENSION = "not_extension"
    """Symbol indicating an :class:`InspectionAttr` that's
    not part of sqlalchemy.ext.

    Is assigned to the :attr:`.InspectionAttr.extension_type`
    attribute.

    """


_never_set = frozenset([NEVER_SET])

_none_set = frozenset([None, NEVER_SET, PASSIVE_NO_RESULT])

_none_only_set = frozenset([None])

_SET_DEFERRED_EXPIRED = util.symbol("SET_DEFERRED_EXPIRED")

_DEFER_FOR_STATE = util.symbol("DEFER_FOR_STATE")

_RAISE_FOR_STATE = util.symbol("RAISE_FOR_STATE")


_F = TypeVar("_F", bound=Callable[..., Any])
_Self = TypeVar("_Self")


def _assertions(
    *assertions: Any,
) -> Callable[[_F], _F]:
    @util.decorator
    def generate(fn: _F, self: _Self, *args: Any, **kw: Any) -> _Self:
        for assertion in assertions:
            assertion(self, fn.__name__)
        fn(self, *args, **kw)
        return self

    return generate


if TYPE_CHECKING:

    def manager_of_class(cls: Type[_O]) -> ClassManager[_O]: ...

    @overload
    def opt_manager_of_class(cls: AliasedClass[Any]) -> None: ...

    @overload
    def opt_manager_of_class(
        cls: _ExternalEntityType[_O],
    ) -> Optional[ClassManager[_O]]: ...

    def opt_manager_of_class(
        cls: _ExternalEntityType[_O],
    ) -> Optional[ClassManager[_O]]: ...

    def instance_state(instance: _O) -> InstanceState[_O]: ...

    def instance_dict(instance: object) -> Dict[str, Any]: ...

else:
    # these can be replaced by sqlalchemy.ext.instrumentation
    # if augmented class instrumentation is enabled.

    def manager_of_class(cls):
        try:
            return cls.__dict__[DEFAULT_MANAGER_ATTR]
        except KeyError as ke:
            raise exc.UnmappedClassError(
                cls, f"Can't locate an instrumentation manager for class {cls}"
            ) from ke

    def opt_manager_of_class(cls):
        return cls.__dict__.get(DEFAULT_MANAGER_ATTR)

    instance_state = operator.attrgetter(DEFAULT_STATE_ATTR)

    instance_dict = operator.attrgetter("__dict__")


def instance_str(instance: object) -> str:
    """Return a string describing an instance."""

    return state_str(instance_state(instance))


def state_str(state: InstanceState[Any]) -> str:
    """Return a string describing an instance via its InstanceState."""

    if state is None:
        return "None"
    else:
        return "<%s at 0x%x>" % (state.class_.__name__, id(state.obj()))


def state_class_str(state: InstanceState[Any]) -> str:
    """Return a string describing an instance's class via its
    InstanceState.
    """

    if state is None:
        return "None"
    else:
        return "<%s>" % (state.class_.__name__,)


def attribute_str(instance: object, attribute: str) -> str:
    return instance_str(instance) + "." + attribute


def state_attribute_str(state: InstanceState[Any], attribute: str) -> str:
    return state_str(state) + "." + attribute


def object_mapper(instance: _T) -> Mapper[_T]:
    """Given an object, return the primary Mapper associated with the object
    instance.

    Raises :class:`sqlalchemy.orm.exc.UnmappedInstanceError`
    if no mapping is configured.

    This function is available via the inspection system as::

        inspect(instance).mapper

    Using the inspection system will raise
    :class:`sqlalchemy.exc.NoInspectionAvailable` if the instance is
    not part of a mapping.

    """
    return object_state(instance).mapper


def object_state(instance: _T) -> InstanceState[_T]:
    """Given an object, return the :class:`.InstanceState`
    associated with the object.

    Raises :class:`sqlalchemy.orm.exc.UnmappedInstanceError`
    if no mapping is configured.

    Equivalent functionality is available via the :func:`_sa.inspect`
    function as::

        inspect(instance)

    Using the inspection system will raise
    :class:`sqlalchemy.exc.NoInspectionAvailable` if the instance is
    not part of a mapping.

    """
    state = _inspect_mapped_object(instance)
    if state is None:
        raise exc.UnmappedInstanceError(instance)
    else:
        return state


@inspection._inspects(object)
def _inspect_mapped_object(instance: _T) -> Optional[InstanceState[_T]]:
    try:
        return instance_state(instance)
    except (exc.UnmappedClassError,) + exc.NO_STATE:
        return None


def _class_to_mapper(
    class_or_mapper: Union[Mapper[_T], Type[_T]]
) -> Mapper[_T]:
    # can't get mypy to see an overload for this
    insp = inspection.inspect(class_or_mapper, False)
    if insp is not None:
        return insp.mapper  # type: ignore
    else:
        assert isinstance(class_or_mapper, type)
        raise exc.UnmappedClassError(class_or_mapper)


def _mapper_or_none(
    entity: Union[Type[_T], _InternalEntityType[_T]]
) -> Optional[Mapper[_T]]:
    """Return the :class:`_orm.Mapper` for the given class or None if the
    class is not mapped.
    """

    # can't get mypy to see an overload for this
    insp = inspection.inspect(entity, False)
    if insp is not None:
        return insp.mapper  # type: ignore
    else:
        return None


def _is_mapped_class(entity: Any) -> bool:
    """Return True if the given object is a mapped class,
    :class:`_orm.Mapper`, or :class:`.AliasedClass`.
    """

    insp = inspection.inspect(entity, False)
    return (
        insp is not None
        and not insp.is_clause_element
        and (insp.is_mapper or insp.is_aliased_class)
    )


def _is_aliased_class(entity: Any) -> bool:
    insp = inspection.inspect(entity, False)
    return insp is not None and getattr(insp, "is_aliased_class", False)


@no_type_check
def _entity_descriptor(entity: _EntityType[Any], key: str) -> Any:
    """Return a class attribute given an entity and string name.

    May return :class:`.InstrumentedAttribute` or user-defined
    attribute.

    """
    insp = inspection.inspect(entity)
    if insp.is_selectable:
        description = entity
        entity = insp.c
    elif insp.is_aliased_class:
        entity = insp.entity
        description = entity
    elif hasattr(insp, "mapper"):
        description = entity = insp.mapper.class_
    else:
        description = entity

    try:
        return getattr(entity, key)
    except AttributeError as err:
        raise sa_exc.InvalidRequestError(
            "Entity '%s' has no property '%s'" % (description, key)
        ) from err


if TYPE_CHECKING:

    def _state_mapper(state: InstanceState[_O]) -> Mapper[_O]: ...

else:
    _state_mapper = util.dottedgetter("manager.mapper")


def _inspect_mapped_class(
    class_: Type[_O], configure: bool = False
) -> Optional[Mapper[_O]]:
    try:
        class_manager = opt_manager_of_class(class_)
        if class_manager is None or not class_manager.is_mapped:
            return None
        mapper = class_manager.mapper
    except exc.NO_STATE:
        return None
    else:
        if configure:
            mapper._check_configure()
        return mapper


def _parse_mapper_argument(arg: Union[Mapper[_O], Type[_O]]) -> Mapper[_O]:
    insp = inspection.inspect(arg, raiseerr=False)
    if insp_is_mapper(insp):
        return insp

    raise sa_exc.ArgumentError(f"Mapper or mapped class expected, got {arg!r}")


def class_mapper(class_: Type[_O], configure: bool = True) -> Mapper[_O]:
    """Given a class, return the primary :class:`_orm.Mapper` associated
    with the key.

    Raises :exc:`.UnmappedClassError` if no mapping is configured
    on the given class, or :exc:`.ArgumentError` if a non-class
    object is passed.

    Equivalent functionality is available via the :func:`_sa.inspect`
    function as::

        inspect(some_mapped_class)

    Using the inspection system will raise
    :class:`sqlalchemy.exc.NoInspectionAvailable` if the class is not mapped.

    """
    mapper = _inspect_mapped_class(class_, configure=configure)
    if mapper is None:
        if not isinstance(class_, type):
            raise sa_exc.ArgumentError(
                "Class object expected, got '%r'." % (class_,)
            )
        raise exc.UnmappedClassError(class_)
    else:
        return mapper


class InspectionAttr:
    """A base class applied to all ORM objects and attributes that are
    related to things that can be returned by the :func:`_sa.inspect` function.

    The attributes defined here allow the usage of simple boolean
    checks to test basic facts about the object returned.

    While the boolean checks here are basically the same as using
    the Python isinstance() function, the flags here can be used without
    the need to import all of these classes, and also such that
    the SQLAlchemy class system can change while leaving the flags
    here intact for forwards-compatibility.

    """

    __slots__: Tuple[str, ...] = ()

    is_selectable = False
    """Return True if this object is an instance of
    :class:`_expression.Selectable`."""

    is_aliased_class = False
    """True if this object is an instance of :class:`.AliasedClass`."""

    is_instance = False
    """True if this object is an instance of :class:`.InstanceState`."""

    is_mapper = False
    """True if this object is an instance of :class:`_orm.Mapper`."""

    is_bundle = False
    """True if this object is an instance of :class:`.Bundle`."""

    is_property = False
    """True if this object is an instance of :class:`.MapperProperty`."""

    is_attribute = False
    """True if this object is a Python :term:`descriptor`.

    This can refer to one of many types.   Usually a
    :class:`.QueryableAttribute` which handles attributes events on behalf
    of a :class:`.MapperProperty`.   But can also be an extension type
    such as :class:`.AssociationProxy` or :class:`.hybrid_property`.
    The :attr:`.InspectionAttr.extension_type` will refer to a constant
    identifying the specific subtype.

    .. seealso::

        :attr:`_orm.Mapper.all_orm_descriptors`

    """

    _is_internal_proxy = False
    """True if this object is an internal proxy object.

    .. versionadded:: 1.2.12

    """

    is_clause_element = False
    """True if this object is an instance of
    :class:`_expression.ClauseElement`."""

    extension_type: InspectionAttrExtensionType = NotExtension.NOT_EXTENSION
    """The extension type, if any.
    Defaults to :attr:`.interfaces.NotExtension.NOT_EXTENSION`

    .. seealso::

        :class:`.HybridExtensionType`

        :class:`.AssociationProxyExtensionType`

    """


class InspectionAttrInfo(InspectionAttr):
    """Adds the ``.info`` attribute to :class:`.InspectionAttr`.

    The rationale for :class:`.InspectionAttr` vs. :class:`.InspectionAttrInfo`
    is that the former is compatible as a mixin for classes that specify
    ``__slots__``; this is essentially an implementation artifact.

    """

    __slots__ = ()

    @util.ro_memoized_property
    def info(self) -> _InfoType:
        """Info dictionary associated with the object, allowing user-defined
        data to be associated with this :class:`.InspectionAttr`.

        The dictionary is generated when first accessed.  Alternatively,
        it can be specified as a constructor argument to the
        :func:`.column_property`, :func:`_orm.relationship`, or
        :func:`.composite`
        functions.

        .. seealso::

            :attr:`.QueryableAttribute.info`

            :attr:`.SchemaItem.info`

        """
        return {}


class SQLORMOperations(SQLCoreOperations[_T_co], TypingOnly):
    __slots__ = ()

    if typing.TYPE_CHECKING:

        def of_type(
            self, class_: _EntityType[Any]
        ) -> PropComparator[_T_co]: ...

        def and_(
            self, *criteria: _ColumnExpressionArgument[bool]
        ) -> PropComparator[bool]: ...

        def any(  # noqa: A001
            self,
            criterion: Optional[_ColumnExpressionArgument[bool]] = None,
            **kwargs: Any,
        ) -> ColumnElement[bool]: ...

        def has(
            self,
            criterion: Optional[_ColumnExpressionArgument[bool]] = None,
            **kwargs: Any,
        ) -> ColumnElement[bool]: ...


class ORMDescriptor(Generic[_T_co], TypingOnly):
    """Represent any Python descriptor that provides a SQL expression
    construct at the class level."""

    __slots__ = ()

    if typing.TYPE_CHECKING:

        @overload
        def __get__(
            self, instance: Any, owner: Literal[None]
        ) -> ORMDescriptor[_T_co]: ...

        @overload
        def __get__(
            self, instance: Literal[None], owner: Any
        ) -> SQLCoreOperations[_T_co]: ...

        @overload
        def __get__(self, instance: object, owner: Any) -> _T_co: ...

        def __get__(
            self, instance: object, owner: Any
        ) -> Union[ORMDescriptor[_T_co], SQLCoreOperations[_T_co], _T_co]: ...


class _MappedAnnotationBase(Generic[_T_co], TypingOnly):
    """common class for Mapped and similar ORM container classes.

    these are classes that can appear on the left side of an ORM declarative
    mapping, containing a mapped class or in some cases a collection
    surrounding a mapped class.

    """

    __slots__ = ()


class SQLORMExpression(
    SQLORMOperations[_T_co], SQLColumnExpression[_T_co], TypingOnly
):
    """A type that may be used to indicate any ORM-level attribute or
    object that acts in place of one, in the context of SQL expression
    construction.

    :class:`.SQLORMExpression` extends from the Core
    :class:`.SQLColumnExpression` to add additional SQL methods that are ORM
    specific, such as :meth:`.PropComparator.of_type`, and is part of the bases
    for :class:`.InstrumentedAttribute`. It may be used in :pep:`484` typing to
    indicate arguments or return values that should behave as ORM-level
    attribute expressions.

    .. versionadded:: 2.0.0b4


    """

    __slots__ = ()


class Mapped(
    SQLORMExpression[_T_co],
    ORMDescriptor[_T_co],
    _MappedAnnotationBase[_T_co],
    roles.DDLConstraintColumnRole,
):
    """Represent an ORM mapped attribute on a mapped class.

    This class represents the complete descriptor interface for any class
    attribute that will have been :term:`instrumented` by the ORM
    :class:`_orm.Mapper` class.   Provides appropriate information to type
    checkers such as pylance and mypy so that ORM-mapped attributes
    are correctly typed.

    The most prominent use of :class:`_orm.Mapped` is in
    the :ref:`Declarative Mapping <orm_explicit_declarative_base>` form
    of :class:`_orm.Mapper` configuration, where used explicitly it drives
    the configuration of ORM attributes such as :func:`_orm.mapped_class`
    and :func:`_orm.relationship`.

    .. seealso::

        :ref:`orm_explicit_declarative_base`

        :ref:`orm_declarative_table`

    .. tip::

        The :class:`_orm.Mapped` class represents attributes that are handled
        directly by the :class:`_orm.Mapper` class. It does not include other
        Python descriptor classes that are provided as extensions, including
        :ref:`hybrids_toplevel` and the :ref:`associationproxy_toplevel`.
        While these systems still make use of ORM-specific superclasses
        and structures, they are not :term:`instrumented` by the
        :class:`_orm.Mapper` and instead provide their own functionality
        when they are accessed on a class.

    .. versionadded:: 1.4


    """

    __slots__ = ()

    if typing.TYPE_CHECKING:

        @overload
        def __get__(
            self, instance: None, owner: Any
        ) -> InstrumentedAttribute[_T_co]: ...

        @overload
        def __get__(self, instance: object, owner: Any) -> _T_co: ...

        def __get__(
            self, instance: Optional[object], owner: Any
        ) -> Union[InstrumentedAttribute[_T_co], _T_co]: ...

        @classmethod
        def _empty_constructor(cls, arg1: Any) -> Mapped[_T_co]: ...

        def __set__(
            self, instance: Any, value: Union[SQLCoreOperations[_T_co], _T_co]
        ) -> None: ...

        def __delete__(self, instance: Any) -> None: ...


class _MappedAttribute(Generic[_T_co], TypingOnly):
    """Mixin for attributes which should be replaced by mapper-assigned
    attributes.

    """

    __slots__ = ()


class _DeclarativeMapped(Mapped[_T_co], _MappedAttribute[_T_co]):
    """Mixin for :class:`.MapperProperty` subclasses that allows them to
    be compatible with ORM-annotated declarative mappings.

    """

    __slots__ = ()

    # MappedSQLExpression, Relationship, Composite etc. dont actually do
    # SQL expression behavior.  yet there is code that compares them with
    # __eq__(), __ne__(), etc.   Since #8847 made Mapped even more full
    # featured including ColumnOperators, we need to have those methods
    # be no-ops for these objects, so return NotImplemented to fall back
    # to normal comparison behavior.
    def operate(self, op: OperatorType, *other: Any, **kwargs: Any) -> Any:
        return NotImplemented

    __sa_operate__ = operate

    def reverse_operate(
        self, op: OperatorType, other: Any, **kwargs: Any
    ) -> Any:
        return NotImplemented


class DynamicMapped(_MappedAnnotationBase[_T_co]):
    """Represent the ORM mapped attribute type for a "dynamic" relationship.

    The :class:`_orm.DynamicMapped` type annotation may be used in an
    :ref:`Annotated Declarative Table <orm_declarative_mapped_column>` mapping
    to indicate that the ``lazy="dynamic"`` loader strategy should be used
    for a particular :func:`_orm.relationship`.

    .. legacy::  The "dynamic" lazy loader strategy is the legacy form of what
       is now the "write_only" strategy described in the section
       :ref:`write_only_relationship`.

    E.g.::

        class User(Base):
            __tablename__ = "user"
            id: Mapped[int] = mapped_column(primary_key=True)
            addresses: DynamicMapped[Address] = relationship(
                cascade="all,delete-orphan"
            )

    See the section :ref:`dynamic_relationship` for background.

    .. versionadded:: 2.0

    .. seealso::

        :ref:`dynamic_relationship` - complete background

        :class:`.WriteOnlyMapped` - fully 2.0 style version

    """

    __slots__ = ()

    if TYPE_CHECKING:

        @overload
        def __get__(
            self, instance: None, owner: Any
        ) -> InstrumentedAttribute[_T_co]: ...

        @overload
        def __get__(
            self, instance: object, owner: Any
        ) -> AppenderQuery[_T_co]: ...

        def __get__(
            self, instance: Optional[object], owner: Any
        ) -> Union[InstrumentedAttribute[_T_co], AppenderQuery[_T_co]]: ...

        def __set__(
            self, instance: Any, value: typing.Collection[_T_co]
        ) -> None: ...


class WriteOnlyMapped(_MappedAnnotationBase[_T_co]):
    """Represent the ORM mapped attribute type for a "write only" relationship.

    The :class:`_orm.WriteOnlyMapped` type annotation may be used in an
    :ref:`Annotated Declarative Table <orm_declarative_mapped_column>` mapping
    to indicate that the ``lazy="write_only"`` loader strategy should be used
    for a particular :func:`_orm.relationship`.

    E.g.::

        class User(Base):
            __tablename__ = "user"
            id: Mapped[int] = mapped_column(primary_key=True)
            addresses: WriteOnlyMapped[Address] = relationship(
                cascade="all,delete-orphan"
            )

    See the section :ref:`write_only_relationship` for background.

    .. versionadded:: 2.0

    .. seealso::

        :ref:`write_only_relationship` - complete background

        :class:`.DynamicMapped` - includes legacy :class:`_orm.Query` support

    """

    __slots__ = ()

    if TYPE_CHECKING:

        @overload
        def __get__(
            self, instance: None, owner: Any
        ) -> InstrumentedAttribute[_T_co]: ...

        @overload
        def __get__(
            self, instance: object, owner: Any
        ) -> WriteOnlyCollection[_T_co]: ...

        def __get__(
            self, instance: Optional[object], owner: Any
        ) -> Union[
            InstrumentedAttribute[_T_co], WriteOnlyCollection[_T_co]
        ]: ...

        def __set__(
            self, instance: Any, value: typing.Collection[_T_co]
        ) -> None: ...

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\cffi\api.py ===
import sys, types
from .lock import allocate_lock
from .error import CDefError
from . import model

try:
    callable
except NameError:
    # Python 3.1
    from collections import Callable
    callable = lambda x: isinstance(x, Callable)

try:
    basestring
except NameError:
    # Python 3.x
    basestring = str

_unspecified = object()



class FFI(object):
    r'''
    The main top-level class that you instantiate once, or once per module.

    Example usage:

        ffi = FFI()
        ffi.cdef("""
            int printf(const char *, ...);
        """)

        C = ffi.dlopen(None)   # standard library
        -or-
        C = ffi.verify()  # use a C compiler: verify the decl above is right

        C.printf("hello, %s!\n", ffi.new("char[]", "world"))
    '''

    def __init__(self, backend=None):
        """Create an FFI instance.  The 'backend' argument is used to
        select a non-default backend, mostly for tests.
        """
        if backend is None:
            # You need PyPy (>= 2.0 beta), or a CPython (>= 2.6) with
            # _cffi_backend.so compiled.
            import _cffi_backend as backend
            from . import __version__
            if backend.__version__ != __version__:
                # bad version!  Try to be as explicit as possible.
                if hasattr(backend, '__file__'):
                    # CPython
                    raise Exception("Version mismatch: this is the 'cffi' package version %s, located in %r.  When we import the top-level '_cffi_backend' extension module, we get version %s, located in %r.  The two versions should be equal; check your installation." % (
                        __version__, __file__,
                        backend.__version__, backend.__file__))
                else:
                    # PyPy
                    raise Exception("Version mismatch: this is the 'cffi' package version %s, located in %r.  This interpreter comes with a built-in '_cffi_backend' module, which is version %s.  The two versions should be equal; check your installation." % (
                        __version__, __file__, backend.__version__))
            # (If you insist you can also try to pass the option
            # 'backend=backend_ctypes.CTypesBackend()', but don't
            # rely on it!  It's probably not going to work well.)

        from . import cparser
        self._backend = backend
        self._lock = allocate_lock()
        self._parser = cparser.Parser()
        self._cached_btypes = {}
        self._parsed_types = types.ModuleType('parsed_types').__dict__
        self._new_types = types.ModuleType('new_types').__dict__
        self._function_caches = []
        self._libraries = []
        self._cdefsources = []
        self._included_ffis = []
        self._windows_unicode = None
        self._init_once_cache = {}
        self._cdef_version = None
        self._embedding = None
        self._typecache = model.get_typecache(backend)
        if hasattr(backend, 'set_ffi'):
            backend.set_ffi(self)
        for name in list(backend.__dict__):
            if name.startswith('RTLD_'):
                setattr(self, name, getattr(backend, name))
        #
        with self._lock:
            self.BVoidP = self._get_cached_btype(model.voidp_type)
            self.BCharA = self._get_cached_btype(model.char_array_type)
        if isinstance(backend, types.ModuleType):
            # _cffi_backend: attach these constants to the class
            if not hasattr(FFI, 'NULL'):
                FFI.NULL = self.cast(self.BVoidP, 0)
                FFI.CData, FFI.CType = backend._get_types()
        else:
            # ctypes backend: attach these constants to the instance
            self.NULL = self.cast(self.BVoidP, 0)
            self.CData, self.CType = backend._get_types()
        self.buffer = backend.buffer

    def cdef(self, csource, override=False, packed=False, pack=None):
        """Parse the given C source.  This registers all declared functions,
        types, and global variables.  The functions and global variables can
        then be accessed via either 'ffi.dlopen()' or 'ffi.verify()'.
        The types can be used in 'ffi.new()' and other functions.
        If 'packed' is specified as True, all structs declared inside this
        cdef are packed, i.e. laid out without any field alignment at all.
        Alternatively, 'pack' can be a small integer, and requests for
        alignment greater than that are ignored (pack=1 is equivalent to
        packed=True).
        """
        self._cdef(csource, override=override, packed=packed, pack=pack)

    def embedding_api(self, csource, packed=False, pack=None):
        self._cdef(csource, packed=packed, pack=pack, dllexport=True)
        if self._embedding is None:
            self._embedding = ''

    def _cdef(self, csource, override=False, **options):
        if not isinstance(csource, str):    # unicode, on Python 2
            if not isinstance(csource, basestring):
                raise TypeError("cdef() argument must be a string")
            csource = csource.encode('ascii')
        with self._lock:
            self._cdef_version = object()
            self._parser.parse(csource, override=override, **options)
            self._cdefsources.append(csource)
            if override:
                for cache in self._function_caches:
                    cache.clear()
            finishlist = self._parser._recomplete
            if finishlist:
                self._parser._recomplete = []
                for tp in finishlist:
                    tp.finish_backend_type(self, finishlist)

    def dlopen(self, name, flags=0):
        """Load and return a dynamic library identified by 'name'.
        The standard C library can be loaded by passing None.
        Note that functions and types declared by 'ffi.cdef()' are not
        linked to a particular library, just like C headers; in the
        library we only look for the actual (untyped) symbols.
        """
        if not (isinstance(name, basestring) or
                name is None or
                isinstance(name, self.CData)):
            raise TypeError("dlopen(name): name must be a file name, None, "
                            "or an already-opened 'void *' handle")
        with self._lock:
            lib, function_cache = _make_ffi_library(self, name, flags)
            self._function_caches.append(function_cache)
            self._libraries.append(lib)
        return lib

    def dlclose(self, lib):
        """Close a library obtained with ffi.dlopen().  After this call,
        access to functions or variables from the library will fail
        (possibly with a segmentation fault).
        """
        type(lib).__cffi_close__(lib)

    def _typeof_locked(self, cdecl):
        # call me with the lock!
        key = cdecl
        if key in self._parsed_types:
            return self._parsed_types[key]
        #
        if not isinstance(cdecl, str):    # unicode, on Python 2
            cdecl = cdecl.encode('ascii')
        #
        type = self._parser.parse_type(cdecl)
        really_a_function_type = type.is_raw_function
        if really_a_function_type:
            type = type.as_function_pointer()
        btype = self._get_cached_btype(type)
        result = btype, really_a_function_type
        self._parsed_types[key] = result
        return result

    def _typeof(self, cdecl, consider_function_as_funcptr=False):
        # string -> ctype object
        try:
            result = self._parsed_types[cdecl]
        except KeyError:
            with self._lock:
                result = self._typeof_locked(cdecl)
        #
        btype, really_a_function_type = result
        if really_a_function_type and not consider_function_as_funcptr:
            raise CDefError("the type %r is a function type, not a "
                            "pointer-to-function type" % (cdecl,))
        return btype

    def typeof(self, cdecl):
        """Parse the C type given as a string and return the
        corresponding <ctype> object.
        It can also be used on 'cdata' instance to get its C type.
        """
        if isinstance(cdecl, basestring):
            return self._typeof(cdecl)
        if isinstance(cdecl, self.CData):
            return self._backend.typeof(cdecl)
        if isinstance(cdecl, types.BuiltinFunctionType):
            res = _builtin_function_type(cdecl)
            if res is not None:
                return res
        if (isinstance(cdecl, types.FunctionType)
                and hasattr(cdecl, '_cffi_base_type')):
            with self._lock:
                return self._get_cached_btype(cdecl._cffi_base_type)
        raise TypeError(type(cdecl))

    def sizeof(self, cdecl):
        """Return the size in bytes of the argument.  It can be a
        string naming a C type, or a 'cdata' instance.
        """
        if isinstance(cdecl, basestring):
            BType = self._typeof(cdecl)
            return self._backend.sizeof(BType)
        else:
            return self._backend.sizeof(cdecl)

    def alignof(self, cdecl):
        """Return the natural alignment size in bytes of the C type
        given as a string.
        """
        if isinstance(cdecl, basestring):
            cdecl = self._typeof(cdecl)
        return self._backend.alignof(cdecl)

    def offsetof(self, cdecl, *fields_or_indexes):
        """Return the offset of the named field inside the given
        structure or array, which must be given as a C type name.
        You can give several field names in case of nested structures.
        You can also give numeric values which correspond to array
        items, in case of an array type.
        """
        if isinstance(cdecl, basestring):
            cdecl = self._typeof(cdecl)
        return self._typeoffsetof(cdecl, *fields_or_indexes)[1]

    def new(self, cdecl, init=None):
        """Allocate an instance according to the specified C type and
        return a pointer to it.  The specified C type must be either a
        pointer or an array: ``new('X *')`` allocates an X and returns
        a pointer to it, whereas ``new('X[n]')`` allocates an array of
        n X'es and returns an array referencing it (which works
        mostly like a pointer, like in C).  You can also use
        ``new('X[]', n)`` to allocate an array of a non-constant
        length n.

        The memory is initialized following the rules of declaring a
        global variable in C: by default it is zero-initialized, but
        an explicit initializer can be given which can be used to
        fill all or part of the memory.

        When the returned <cdata> object goes out of scope, the memory
        is freed.  In other words the returned <cdata> object has
        ownership of the value of type 'cdecl' that it points to.  This
        means that the raw data can be used as long as this object is
        kept alive, but must not be used for a longer time.  Be careful
        about that when copying the pointer to the memory somewhere
        else, e.g. into another structure.
        """
        if isinstance(cdecl, basestring):
            cdecl = self._typeof(cdecl)
        return self._backend.newp(cdecl, init)

    def new_allocator(self, alloc=None, free=None,
                      should_clear_after_alloc=True):
        """Return a new allocator, i.e. a function that behaves like ffi.new()
        but uses the provided low-level 'alloc' and 'free' functions.

        'alloc' is called with the size as argument.  If it returns NULL, a
        MemoryError is raised.  'free' is called with the result of 'alloc'
        as argument.  Both can be either Python function or directly C
        functions.  If 'free' is None, then no free function is called.
        If both 'alloc' and 'free' are None, the default is used.

        If 'should_clear_after_alloc' is set to False, then the memory
        returned by 'alloc' is assumed to be already cleared (or you are
        fine with garbage); otherwise CFFI will clear it.
        """
        compiled_ffi = self._backend.FFI()
        allocator = compiled_ffi.new_allocator(alloc, free,
                                               should_clear_after_alloc)
        def allocate(cdecl, init=None):
            if isinstance(cdecl, basestring):
                cdecl = self._typeof(cdecl)
            return allocator(cdecl, init)
        return allocate

    def cast(self, cdecl, source):
        """Similar to a C cast: returns an instance of the named C
        type initialized with the given 'source'.  The source is
        casted between integers or pointers of any type.
        """
        if isinstance(cdecl, basestring):
            cdecl = self._typeof(cdecl)
        return self._backend.cast(cdecl, source)

    def string(self, cdata, maxlen=-1):
        """Return a Python string (or unicode string) from the 'cdata'.
        If 'cdata' is a pointer or array of characters or bytes, returns
        the null-terminated string.  The returned string extends until
        the first null character, or at most 'maxlen' characters.  If
        'cdata' is an array then 'maxlen' defaults to its length.

        If 'cdata' is a pointer or array of wchar_t, returns a unicode
        string following the same rules.

        If 'cdata' is a single character or byte or a wchar_t, returns
        it as a string or unicode string.

        If 'cdata' is an enum, returns the value of the enumerator as a
        string, or 'NUMBER' if the value is out of range.
        """
        return self._backend.string(cdata, maxlen)

    def unpack(self, cdata, length):
        """Unpack an array of C data of the given length,
        returning a Python string/unicode/list.

        If 'cdata' is a pointer to 'char', returns a byte string.
        It does not stop at the first null.  This is equivalent to:
        ffi.buffer(cdata, length)[:]

        If 'cdata' is a pointer to 'wchar_t', returns a unicode string.
        'length' is measured in wchar_t's; it is not the size in bytes.

        If 'cdata' is a pointer to anything else, returns a list of
        'length' items.  This is a faster equivalent to:
        [cdata[i] for i in range(length)]
        """
        return self._backend.unpack(cdata, length)

   #def buffer(self, cdata, size=-1):
   #    """Return a read-write buffer object that references the raw C data
   #    pointed to by the given 'cdata'.  The 'cdata' must be a pointer or
   #    an array.  Can be passed to functions expecting a buffer, or directly
   #    manipulated with:
   #
   #        buf[:]          get a copy of it in a regular string, or
   #        buf[idx]        as a single character
   #        buf[:] = ...
   #        buf[idx] = ...  change the content
   #    """
   #    note that 'buffer' is a type, set on this instance by __init__

    def from_buffer(self, cdecl, python_buffer=_unspecified,
                    require_writable=False):
        """Return a cdata of the given type pointing to the data of the
        given Python object, which must support the buffer interface.
        Note that this is not meant to be used on the built-in types
        str or unicode (you can build 'char[]' arrays explicitly)
        but only on objects containing large quantities of raw data
        in some other format, like 'array.array' or numpy arrays.

        The first argument is optional and default to 'char[]'.
        """
        if python_buffer is _unspecified:
            cdecl, python_buffer = self.BCharA, cdecl
        elif isinstance(cdecl, basestring):
            cdecl = self._typeof(cdecl)
        return self._backend.from_buffer(cdecl, python_buffer,
                                         require_writable)

    def memmove(self, dest, src, n):
        """ffi.memmove(dest, src, n) copies n bytes of memory from src to dest.

        Like the C function memmove(), the memory areas may overlap;
        apart from that it behaves like the C function memcpy().

        'src' can be any cdata ptr or array, or any Python buffer object.
        'dest' can be any cdata ptr or array, or a writable Python buffer
        object.  The size to copy, 'n', is always measured in bytes.

        Unlike other methods, this one supports all Python buffer including
        byte strings and bytearrays---but it still does not support
        non-contiguous buffers.
        """
        return self._backend.memmove(dest, src, n)

    def callback(self, cdecl, python_callable=None, error=None, onerror=None):
        """Return a callback object or a decorator making such a
        callback object.  'cdecl' must name a C function pointer type.
        The callback invokes the specified 'python_callable' (which may
        be provided either directly or via a decorator).  Important: the
        callback object must be manually kept alive for as long as the
        callback may be invoked from the C level.
        """
        def callback_decorator_wrap(python_callable):
            if not callable(python_callable):
                raise TypeError("the 'python_callable' argument "
                                "is not callable")
            return self._backend.callback(cdecl, python_callable,
                                          error, onerror)
        if isinstance(cdecl, basestring):
            cdecl = self._typeof(cdecl, consider_function_as_funcptr=True)
        if python_callable is None:
            return callback_decorator_wrap                # decorator mode
        else:
            return callback_decorator_wrap(python_callable)  # direct mode

    def getctype(self, cdecl, replace_with=''):
        """Return a string giving the C type 'cdecl', which may be itself
        a string or a <ctype> object.  If 'replace_with' is given, it gives
        extra text to append (or insert for more complicated C types), like
        a variable name, or '*' to get actually the C type 'pointer-to-cdecl'.
        """
        if isinstance(cdecl, basestring):
            cdecl = self._typeof(cdecl)
        replace_with = replace_with.strip()
        if (replace_with.startswith('*')
                and '&[' in self._backend.getcname(cdecl, '&')):
            replace_with = '(%s)' % replace_with
        elif replace_with and not replace_with[0] in '[(':
            replace_with = ' ' + replace_with
        return self._backend.getcname(cdecl, replace_with)

    def gc(self, cdata, destructor, size=0):
        """Return a new cdata object that points to the same
        data.  Later, when this new cdata object is garbage-collected,
        'destructor(old_cdata_object)' will be called.

        The optional 'size' gives an estimate of the size, used to
        trigger the garbage collection more eagerly.  So far only used
        on PyPy.  It tells the GC that the returned object keeps alive
        roughly 'size' bytes of external memory.
        """
        return self._backend.gcp(cdata, destructor, size)

    def _get_cached_btype(self, type):
        assert self._lock.acquire(False) is False
        # call me with the lock!
        try:
            BType = self._cached_btypes[type]
        except KeyError:
            finishlist = []
            BType = type.get_cached_btype(self, finishlist)
            for type in finishlist:
                type.finish_backend_type(self, finishlist)
        return BType

    def verify(self, source='', tmpdir=None, **kwargs):
        """Verify that the current ffi signatures compile on this
        machine, and return a dynamic library object.  The dynamic
        library can be used to call functions and access global
        variables declared in this 'ffi'.  The library is compiled
        by the C compiler: it gives you C-level API compatibility
        (including calling macros).  This is unlike 'ffi.dlopen()',
        which requires binary compatibility in the signatures.
        """
        from .verifier import Verifier, _caller_dir_pycache
        #
        # If set_unicode(True) was called, insert the UNICODE and
        # _UNICODE macro declarations
        if self._windows_unicode:
            self._apply_windows_unicode(kwargs)
        #
        # Set the tmpdir here, and not in Verifier.__init__: it picks
        # up the caller's directory, which we want to be the caller of
        # ffi.verify(), as opposed to the caller of Veritier().
        tmpdir = tmpdir or _caller_dir_pycache()
        #
        # Make a Verifier() and use it to load the library.
        self.verifier = Verifier(self, source, tmpdir, **kwargs)
        lib = self.verifier.load_library()
        #
        # Save the loaded library for keep-alive purposes, even
        # if the caller doesn't keep it alive itself (it should).
        self._libraries.append(lib)
        return lib

    def _get_errno(self):
        return self._backend.get_errno()
    def _set_errno(self, errno):
        self._backend.set_errno(errno)
    errno = property(_get_errno, _set_errno, None,
                     "the value of 'errno' from/to the C calls")

    def getwinerror(self, code=-1):
        return self._backend.getwinerror(code)

    def _pointer_to(self, ctype):
        with self._lock:
            return model.pointer_cache(self, ctype)

    def addressof(self, cdata, *fields_or_indexes):
        """Return the address of a <cdata 'struct-or-union'>.
        If 'fields_or_indexes' are given, returns the address of that
        field or array item in the structure or array, recursively in
        case of nested structures.
        """
        try:
            ctype = self._backend.typeof(cdata)
        except TypeError:
            if '__addressof__' in type(cdata).__dict__:
                return type(cdata).__addressof__(cdata, *fields_or_indexes)
            raise
        if fields_or_indexes:
            ctype, offset = self._typeoffsetof(ctype, *fields_or_indexes)
        else:
            if ctype.kind == "pointer":
                raise TypeError("addressof(pointer)")
            offset = 0
        ctypeptr = self._pointer_to(ctype)
        return self._backend.rawaddressof(ctypeptr, cdata, offset)

    def _typeoffsetof(self, ctype, field_or_index, *fields_or_indexes):
        ctype, offset = self._backend.typeoffsetof(ctype, field_or_index)
        for field1 in fields_or_indexes:
            ctype, offset1 = self._backend.typeoffsetof(ctype, field1, 1)
            offset += offset1
        return ctype, offset

    def include(self, ffi_to_include):
        """Includes the typedefs, structs, unions and enums defined
        in another FFI instance.  Usage is similar to a #include in C,
        where a part of the program might include types defined in
        another part for its own usage.  Note that the include()
        method has no effect on functions, constants and global
        variables, which must anyway be accessed directly from the
        lib object returned by the original FFI instance.
        """
        if not isinstance(ffi_to_include, FFI):
            raise TypeError("ffi.include() expects an argument that is also of"
                            " type cffi.FFI, not %r" % (
                                type(ffi_to_include).__name__,))
        if ffi_to_include is self:
            raise ValueError("self.include(self)")
        with ffi_to_include._lock:
            with self._lock:
                self._parser.include(ffi_to_include._parser)
                self._cdefsources.append('[')
                self._cdefsources.extend(ffi_to_include._cdefsources)
                self._cdefsources.append(']')
                self._included_ffis.append(ffi_to_include)

    def new_handle(self, x):
        return self._backend.newp_handle(self.BVoidP, x)

    def from_handle(self, x):
        return self._backend.from_handle(x)

    def release(self, x):
        self._backend.release(x)

    def set_unicode(self, enabled_flag):
        """Windows: if 'enabled_flag' is True, enable the UNICODE and
        _UNICODE defines in C, and declare the types like TCHAR and LPTCSTR
        to be (pointers to) wchar_t.  If 'enabled_flag' is False,
        declare these types to be (pointers to) plain 8-bit characters.
        This is mostly for backward compatibility; you usually want True.
        """
        if self._windows_unicode is not None:
            raise ValueError("set_unicode() can only be called once")
        enabled_flag = bool(enabled_flag)
        if enabled_flag:
            self.cdef("typedef wchar_t TBYTE;"
                      "typedef wchar_t TCHAR;"
                      "typedef const wchar_t *LPCTSTR;"
                      "typedef const wchar_t *PCTSTR;"
                      "typedef wchar_t *LPTSTR;"
                      "typedef wchar_t *PTSTR;"
                      "typedef TBYTE *PTBYTE;"
                      "typedef TCHAR *PTCHAR;")
        else:
            self.cdef("typedef char TBYTE;"
                      "typedef char TCHAR;"
                      "typedef const char *LPCTSTR;"
                      "typedef const char *PCTSTR;"
                      "typedef char *LPTSTR;"
                      "typedef char *PTSTR;"
                      "typedef TBYTE *PTBYTE;"
                      "typedef TCHAR *PTCHAR;")
        self._windows_unicode = enabled_flag

    def _apply_windows_unicode(self, kwds):
        defmacros = kwds.get('define_macros', ())
        if not isinstance(defmacros, (list, tuple)):
            raise TypeError("'define_macros' must be a list or tuple")
        defmacros = list(defmacros) + [('UNICODE', '1'),
                                       ('_UNICODE', '1')]
        kwds['define_macros'] = defmacros

    def _apply_embedding_fix(self, kwds):
        # must include an argument like "-lpython2.7" for the compiler
        def ensure(key, value):
            lst = kwds.setdefault(key, [])
            if value not in lst:
                lst.append(value)
        #
        if '__pypy__' in sys.builtin_module_names:
            import os
            if sys.platform == "win32":
                # we need 'libpypy-c.lib'.  Current distributions of
                # pypy (>= 4.1) contain it as 'libs/python27.lib'.
                pythonlib = "python{0[0]}{0[1]}".format(sys.version_info)
                if hasattr(sys, 'prefix'):
                    ensure('library_dirs', os.path.join(sys.prefix, 'libs'))
            else:
                # we need 'libpypy-c.{so,dylib}', which should be by
                # default located in 'sys.prefix/bin' for installed
                # systems.
                if sys.version_info < (3,):
                    pythonlib = "pypy-c"
                else:
                    pythonlib = "pypy3-c"
                if hasattr(sys, 'prefix'):
                    ensure('library_dirs', os.path.join(sys.prefix, 'bin'))
            # On uninstalled pypy's, the libpypy-c is typically found in
            # .../pypy/goal/.
            if hasattr(sys, 'prefix'):
                ensure('library_dirs', os.path.join(sys.prefix, 'pypy', 'goal'))
        else:
            if sys.platform == "win32":
                template = "python%d%d"
                if hasattr(sys, 'gettotalrefcount'):
                    template += '_d'
            else:
                try:
                    import sysconfig
                except ImportError:    # 2.6
                    from cffi._shimmed_dist_utils import sysconfig
                template = "python%d.%d"
                if sysconfig.get_config_var('DEBUG_EXT'):
                    template += sysconfig.get_config_var('DEBUG_EXT')
            pythonlib = (template %
                    (sys.hexversion >> 24, (sys.hexversion >> 16) & 0xff))
            if hasattr(sys, 'abiflags'):
                pythonlib += sys.abiflags
        ensure('libraries', pythonlib)
        if sys.platform == "win32":
            ensure('extra_link_args', '/MANIFEST')

    def set_source(self, module_name, source, source_extension='.c', **kwds):
        import os
        if hasattr(self, '_assigned_source'):
            raise ValueError("set_source() cannot be called several times "
                             "per ffi object")
        if not isinstance(module_name, basestring):
            raise TypeError("'module_name' must be a string")
        if os.sep in module_name or (os.altsep and os.altsep in module_name):
            raise ValueError("'module_name' must not contain '/': use a dotted "
                             "name to make a 'package.module' location")
        self._assigned_source = (str(module_name), source,
                                 source_extension, kwds)

    def set_source_pkgconfig(self, module_name, pkgconfig_libs, source,
                             source_extension='.c', **kwds):
        from . import pkgconfig
        if not isinstance(pkgconfig_libs, list):
            raise TypeError("the pkgconfig_libs argument must be a list "
                            "of package names")
        kwds2 = pkgconfig.flags_from_pkgconfig(pkgconfig_libs)
        pkgconfig.merge_flags(kwds, kwds2)
        self.set_source(module_name, source, source_extension, **kwds)

    def distutils_extension(self, tmpdir='build', verbose=True):
        from cffi._shimmed_dist_utils import mkpath
        from .recompiler import recompile
        #
        if not hasattr(self, '_assigned_source'):
            if hasattr(self, 'verifier'):     # fallback, 'tmpdir' ignored
                return self.verifier.get_extension()
            raise ValueError("set_source() must be called before"
                             " distutils_extension()")
        module_name, source, source_extension, kwds = self._assigned_source
        if source is None:
            raise TypeError("distutils_extension() is only for C extension "
                            "modules, not for dlopen()-style pure Python "
                            "modules")
        mkpath(tmpdir)
        ext, updated = recompile(self, module_name,
                                 source, tmpdir=tmpdir, extradir=tmpdir,
                                 source_extension=source_extension,
                                 call_c_compiler=False, **kwds)
        if verbose:
            if updated:
                sys.stderr.write("regenerated: %r\n" % (ext.sources[0],))
            else:
                sys.stderr.write("not modified: %r\n" % (ext.sources[0],))
        return ext

    def emit_c_code(self, filename):
        from .recompiler import recompile
        #
        if not hasattr(self, '_assigned_source'):
            raise ValueError("set_source() must be called before emit_c_code()")
        module_name, source, source_extension, kwds = self._assigned_source
        if source is None:
            raise TypeError("emit_c_code() is only for C extension modules, "
                            "not for dlopen()-style pure Python modules")
        recompile(self, module_name, source,
                  c_file=filename, call_c_compiler=False,
                  uses_ffiplatform=False, **kwds)

    def emit_python_code(self, filename):
        from .recompiler import recompile
        #
        if not hasattr(self, '_assigned_source'):
            raise ValueError("set_source() must be called before emit_c_code()")
        module_name, source, source_extension, kwds = self._assigned_source
        if source is not None:
            raise TypeError("emit_python_code() is only for dlopen()-style "
                            "pure Python modules, not for C extension modules")
        recompile(self, module_name, source,
                  c_file=filename, call_c_compiler=False,
                  uses_ffiplatform=False, **kwds)

    def compile(self, tmpdir='.', verbose=0, target=None, debug=None):
        """The 'target' argument gives the final file name of the
        compiled DLL.  Use '*' to force distutils' choice, suitable for
        regular CPython C API modules.  Use a file name ending in '.*'
        to ask for the system's default extension for dynamic libraries
        (.so/.dll/.dylib).

        The default is '*' when building a non-embedded C API extension,
        and (module_name + '.*') when building an embedded library.
        """
        from .recompiler import recompile
        #
        if not hasattr(self, '_assigned_source'):
            raise ValueError("set_source() must be called before compile()")
        module_name, source, source_extension, kwds = self._assigned_source
        return recompile(self, module_name, source, tmpdir=tmpdir,
                         target=target, source_extension=source_extension,
                         compiler_verbose=verbose, debug=debug, **kwds)

    def init_once(self, func, tag):
        # Read _init_once_cache[tag], which is either (False, lock) if
        # we're calling the function now in some thread, or (True, result).
        # Don't call setdefault() in most cases, to avoid allocating and
        # immediately freeing a lock; but still use setdefaut() to avoid
        # races.
        try:
            x = self._init_once_cache[tag]
        except KeyError:
            x = self._init_once_cache.setdefault(tag, (False, allocate_lock()))
        # Common case: we got (True, result), so we return the result.
        if x[0]:
            return x[1]
        # Else, it's a lock.  Acquire it to serialize the following tests.
        with x[1]:
            # Read again from _init_once_cache the current status.
            x = self._init_once_cache[tag]
            if x[0]:
                return x[1]
            # Call the function and store the result back.
            result = func()
            self._init_once_cache[tag] = (True, result)
        return result

    def embedding_init_code(self, pysource):
        if self._embedding:
            raise ValueError("embedding_init_code() can only be called once")
        # fix 'pysource' before it gets dumped into the C file:
        # - remove empty lines at the beginning, so it starts at "line 1"
        # - dedent, if all non-empty lines are indented
        # - check for SyntaxErrors
        import re
        match = re.match(r'\s*\n', pysource)
        if match:
            pysource = pysource[match.end():]
        lines = pysource.splitlines() or ['']
        prefix = re.match(r'\s*', lines[0]).group()
        for i in range(1, len(lines)):
            line = lines[i]
            if line.rstrip():
                while not line.startswith(prefix):
                    prefix = prefix[:-1]
        i = len(prefix)
        lines = [line[i:]+'\n' for line in lines]
        pysource = ''.join(lines)
        #
        compile(pysource, "cffi_init", "exec")
        #
        self._embedding = pysource

    def def_extern(self, *args, **kwds):
        raise ValueError("ffi.def_extern() is only available on API-mode FFI "
                         "objects")

    def list_types(self):
        """Returns the user type names known to this FFI instance.
        This returns a tuple containing three lists of names:
        (typedef_names, names_of_structs, names_of_unions)
        """
        typedefs = []
        structs = []
        unions = []
        for key in self._parser._declarations:
            if key.startswith('typedef '):
                typedefs.append(key[8:])
            elif key.startswith('struct '):
                structs.append(key[7:])
            elif key.startswith('union '):
                unions.append(key[6:])
        typedefs.sort()
        structs.sort()
        unions.sort()
        return (typedefs, structs, unions)


def _load_backend_lib(backend, name, flags):
    import os
    if not isinstance(name, basestring):
        if sys.platform != "win32" or name is not None:
            return backend.load_library(name, flags)
        name = "c"    # Windows: load_library(None) fails, but this works
                      # on Python 2 (backward compatibility hack only)
    first_error = None
    if '.' in name or '/' in name or os.sep in name:
        try:
            return backend.load_library(name, flags)
        except OSError as e:
            first_error = e
    import ctypes.util
    path = ctypes.util.find_library(name)
    if path is None:
        if name == "c" and sys.platform == "win32" and sys.version_info >= (3,):
            raise OSError("dlopen(None) cannot work on Windows for Python 3 "
                          "(see http://bugs.python.org/issue23606)")
        msg = ("ctypes.util.find_library() did not manage "
               "to locate a library called %r" % (name,))
        if first_error is not None:
            msg = "%s.  Additionally, %s" % (first_error, msg)
        raise OSError(msg)
    return backend.load_library(path, flags)

def _make_ffi_library(ffi, libname, flags):
    backend = ffi._backend
    backendlib = _load_backend_lib(backend, libname, flags)
    #
    def accessor_function(name):
        key = 'function ' + name
        tp, _ = ffi._parser._declarations[key]
        BType = ffi._get_cached_btype(tp)
        value = backendlib.load_function(BType, name)
        library.__dict__[name] = value
    #
    def accessor_variable(name):
        key = 'variable ' + name
        tp, _ = ffi._parser._declarations[key]
        BType = ffi._get_cached_btype(tp)
        read_variable = backendlib.read_variable
        write_variable = backendlib.write_variable
        setattr(FFILibrary, name, property(
            lambda self: read_variable(BType, name),
            lambda self, value: write_variable(BType, name, value)))
    #
    def addressof_var(name):
        try:
            return addr_variables[name]
        except KeyError:
            with ffi._lock:
                if name not in addr_variables:
                    key = 'variable ' + name
                    tp, _ = ffi._parser._declarations[key]
                    BType = ffi._get_cached_btype(tp)
                    if BType.kind != 'array':
                        BType = model.pointer_cache(ffi, BType)
                    p = backendlib.load_function(BType, name)
                    addr_variables[name] = p
            return addr_variables[name]
    #
    def accessor_constant(name):
        raise NotImplementedError("non-integer constant '%s' cannot be "
                                  "accessed from a dlopen() library" % (name,))
    #
    def accessor_int_constant(name):
        library.__dict__[name] = ffi._parser._int_constants[name]
    #
    accessors = {}
    accessors_version = [False]
    addr_variables = {}
    #
    def update_accessors():
        if accessors_version[0] is ffi._cdef_version:
            return
        #
        for key, (tp, _) in ffi._parser._declarations.items():
            if not isinstance(tp, model.EnumType):
                tag, name = key.split(' ', 1)
                if tag == 'function':
                    accessors[name] = accessor_function
                elif tag == 'variable':
                    accessors[name] = accessor_variable
                elif tag == 'constant':
                    accessors[name] = accessor_constant
            else:
                for i, enumname in enumerate(tp.enumerators):
                    def accessor_enum(name, tp=tp, i=i):
                        tp.check_not_partial()
                        library.__dict__[name] = tp.enumvalues[i]
                    accessors[enumname] = accessor_enum
        for name in ffi._parser._int_constants:
            accessors.setdefault(name, accessor_int_constant)
        accessors_version[0] = ffi._cdef_version
    #
    def make_accessor(name):
        with ffi._lock:
            if name in library.__dict__ or name in FFILibrary.__dict__:
                return    # added by another thread while waiting for the lock
            if name not in accessors:
                update_accessors()
                if name not in accessors:
                    raise AttributeError(name)
            accessors[name](name)
    #
    class FFILibrary(object):
        def __getattr__(self, name):
            make_accessor(name)
            return getattr(self, name)
        def __setattr__(self, name, value):
            try:
                property = getattr(self.__class__, name)
            except AttributeError:
                make_accessor(name)
                setattr(self, name, value)
            else:
                property.__set__(self, value)
        def __dir__(self):
            with ffi._lock:
                update_accessors()
                return accessors.keys()
        def __addressof__(self, name):
            if name in library.__dict__:
                return library.__dict__[name]
            if name in FFILibrary.__dict__:
                return addressof_var(name)
            make_accessor(name)
            if name in library.__dict__:
                return library.__dict__[name]
            if name in FFILibrary.__dict__:
                return addressof_var(name)
            raise AttributeError("cffi library has no function or "
                                 "global variable named '%s'" % (name,))
        def __cffi_close__(self):
            backendlib.close_lib()
            self.__dict__.clear()
    #
    if isinstance(libname, basestring):
        try:
            if not isinstance(libname, str):    # unicode, on Python 2
                libname = libname.encode('utf-8')
            FFILibrary.__name__ = 'FFILibrary_%s' % libname
        except UnicodeError:
            pass
    library = FFILibrary()
    return library, library.__dict__

def _builtin_function_type(func):
    # a hack to make at least ffi.typeof(builtin_function) work,
    # if the builtin function was obtained by 'vengine_cpy'.
    import sys
    try:
        module = sys.modules[func.__module__]
        ffi = module._cffi_original_ffi
        types_of_builtin_funcs = module._cffi_types_of_builtin_funcs
        tp = types_of_builtin_funcs[func]
    except (KeyError, AttributeError, TypeError):
        return None
    else:
        with ffi._lock:
            return ffi._get_cached_btype(tp)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\syntax.py ===
import os.path
import re
import sys
import textwrap
from abc import ABC, abstractmethod
from pathlib import Path
from typing import (
    Any,
    Dict,
    Iterable,
    List,
    NamedTuple,
    Optional,
    Sequence,
    Set,
    Tuple,
    Type,
    Union,
)

from pip._vendor.pygments.lexer import Lexer
from pip._vendor.pygments.lexers import get_lexer_by_name, guess_lexer_for_filename
from pip._vendor.pygments.style import Style as PygmentsStyle
from pip._vendor.pygments.styles import get_style_by_name
from pip._vendor.pygments.token import (
    Comment,
    Error,
    Generic,
    Keyword,
    Name,
    Number,
    Operator,
    String,
    Token,
    Whitespace,
)
from pip._vendor.pygments.util import ClassNotFound

from pip._vendor.rich.containers import Lines
from pip._vendor.rich.padding import Padding, PaddingDimensions

from ._loop import loop_first
from .cells import cell_len
from .color import Color, blend_rgb
from .console import Console, ConsoleOptions, JustifyMethod, RenderResult
from .jupyter import JupyterMixin
from .measure import Measurement
from .segment import Segment, Segments
from .style import Style, StyleType
from .text import Text

TokenType = Tuple[str, ...]

WINDOWS = sys.platform == "win32"
DEFAULT_THEME = "monokai"

# The following styles are based on https://github.com/pygments/pygments/blob/master/pygments/formatters/terminal.py
# A few modifications were made

ANSI_LIGHT: Dict[TokenType, Style] = {
    Token: Style(),
    Whitespace: Style(color="white"),
    Comment: Style(dim=True),
    Comment.Preproc: Style(color="cyan"),
    Keyword: Style(color="blue"),
    Keyword.Type: Style(color="cyan"),
    Operator.Word: Style(color="magenta"),
    Name.Builtin: Style(color="cyan"),
    Name.Function: Style(color="green"),
    Name.Namespace: Style(color="cyan", underline=True),
    Name.Class: Style(color="green", underline=True),
    Name.Exception: Style(color="cyan"),
    Name.Decorator: Style(color="magenta", bold=True),
    Name.Variable: Style(color="red"),
    Name.Constant: Style(color="red"),
    Name.Attribute: Style(color="cyan"),
    Name.Tag: Style(color="bright_blue"),
    String: Style(color="yellow"),
    Number: Style(color="blue"),
    Generic.Deleted: Style(color="bright_red"),
    Generic.Inserted: Style(color="green"),
    Generic.Heading: Style(bold=True),
    Generic.Subheading: Style(color="magenta", bold=True),
    Generic.Prompt: Style(bold=True),
    Generic.Error: Style(color="bright_red"),
    Error: Style(color="red", underline=True),
}

ANSI_DARK: Dict[TokenType, Style] = {
    Token: Style(),
    Whitespace: Style(color="bright_black"),
    Comment: Style(dim=True),
    Comment.Preproc: Style(color="bright_cyan"),
    Keyword: Style(color="bright_blue"),
    Keyword.Type: Style(color="bright_cyan"),
    Operator.Word: Style(color="bright_magenta"),
    Name.Builtin: Style(color="bright_cyan"),
    Name.Function: Style(color="bright_green"),
    Name.Namespace: Style(color="bright_cyan", underline=True),
    Name.Class: Style(color="bright_green", underline=True),
    Name.Exception: Style(color="bright_cyan"),
    Name.Decorator: Style(color="bright_magenta", bold=True),
    Name.Variable: Style(color="bright_red"),
    Name.Constant: Style(color="bright_red"),
    Name.Attribute: Style(color="bright_cyan"),
    Name.Tag: Style(color="bright_blue"),
    String: Style(color="yellow"),
    Number: Style(color="bright_blue"),
    Generic.Deleted: Style(color="bright_red"),
    Generic.Inserted: Style(color="bright_green"),
    Generic.Heading: Style(bold=True),
    Generic.Subheading: Style(color="bright_magenta", bold=True),
    Generic.Prompt: Style(bold=True),
    Generic.Error: Style(color="bright_red"),
    Error: Style(color="red", underline=True),
}

RICH_SYNTAX_THEMES = {"ansi_light": ANSI_LIGHT, "ansi_dark": ANSI_DARK}
NUMBERS_COLUMN_DEFAULT_PADDING = 2


class SyntaxTheme(ABC):
    """Base class for a syntax theme."""

    @abstractmethod
    def get_style_for_token(self, token_type: TokenType) -> Style:
        """Get a style for a given Pygments token."""
        raise NotImplementedError  # pragma: no cover

    @abstractmethod
    def get_background_style(self) -> Style:
        """Get the background color."""
        raise NotImplementedError  # pragma: no cover


class PygmentsSyntaxTheme(SyntaxTheme):
    """Syntax theme that delegates to Pygments theme."""

    def __init__(self, theme: Union[str, Type[PygmentsStyle]]) -> None:
        self._style_cache: Dict[TokenType, Style] = {}
        if isinstance(theme, str):
            try:
                self._pygments_style_class = get_style_by_name(theme)
            except ClassNotFound:
                self._pygments_style_class = get_style_by_name("default")
        else:
            self._pygments_style_class = theme

        self._background_color = self._pygments_style_class.background_color
        self._background_style = Style(bgcolor=self._background_color)

    def get_style_for_token(self, token_type: TokenType) -> Style:
        """Get a style from a Pygments class."""
        try:
            return self._style_cache[token_type]
        except KeyError:
            try:
                pygments_style = self._pygments_style_class.style_for_token(token_type)
            except KeyError:
                style = Style.null()
            else:
                color = pygments_style["color"]
                bgcolor = pygments_style["bgcolor"]
                style = Style(
                    color="#" + color if color else "#000000",
                    bgcolor="#" + bgcolor if bgcolor else self._background_color,
                    bold=pygments_style["bold"],
                    italic=pygments_style["italic"],
                    underline=pygments_style["underline"],
                )
            self._style_cache[token_type] = style
        return style

    def get_background_style(self) -> Style:
        return self._background_style


class ANSISyntaxTheme(SyntaxTheme):
    """Syntax theme to use standard colors."""

    def __init__(self, style_map: Dict[TokenType, Style]) -> None:
        self.style_map = style_map
        self._missing_style = Style.null()
        self._background_style = Style.null()
        self._style_cache: Dict[TokenType, Style] = {}

    def get_style_for_token(self, token_type: TokenType) -> Style:
        """Look up style in the style map."""
        try:
            return self._style_cache[token_type]
        except KeyError:
            # Styles form a hierarchy
            # We need to go from most to least specific
            # e.g. ("foo", "bar", "baz") to ("foo", "bar")  to ("foo",)
            get_style = self.style_map.get
            token = tuple(token_type)
            style = self._missing_style
            while token:
                _style = get_style(token)
                if _style is not None:
                    style = _style
                    break
                token = token[:-1]
            self._style_cache[token_type] = style
            return style

    def get_background_style(self) -> Style:
        return self._background_style


SyntaxPosition = Tuple[int, int]


class _SyntaxHighlightRange(NamedTuple):
    """
    A range to highlight in a Syntax object.
    `start` and `end` are 2-integers tuples, where the first integer is the line number
    (starting from 1) and the second integer is the column index (starting from 0).
    """

    style: StyleType
    start: SyntaxPosition
    end: SyntaxPosition
    style_before: bool = False


class Syntax(JupyterMixin):
    """Construct a Syntax object to render syntax highlighted code.

    Args:
        code (str): Code to highlight.
        lexer (Lexer | str): Lexer to use (see https://pygments.org/docs/lexers/)
        theme (str, optional): Color theme, aka Pygments style (see https://pygments.org/docs/styles/#getting-a-list-of-available-styles). Defaults to "monokai".
        dedent (bool, optional): Enable stripping of initial whitespace. Defaults to False.
        line_numbers (bool, optional): Enable rendering of line numbers. Defaults to False.
        start_line (int, optional): Starting number for line numbers. Defaults to 1.
        line_range (Tuple[int | None, int | None], optional): If given should be a tuple of the start and end line to render.
            A value of None in the tuple indicates the range is open in that direction.
        highlight_lines (Set[int]): A set of line numbers to highlight.
        code_width: Width of code to render (not including line numbers), or ``None`` to use all available width.
        tab_size (int, optional): Size of tabs. Defaults to 4.
        word_wrap (bool, optional): Enable word wrapping.
        background_color (str, optional): Optional background color, or None to use theme color. Defaults to None.
        indent_guides (bool, optional): Show indent guides. Defaults to False.
        padding (PaddingDimensions): Padding to apply around the syntax. Defaults to 0 (no padding).
    """

    _pygments_style_class: Type[PygmentsStyle]
    _theme: SyntaxTheme

    @classmethod
    def get_theme(cls, name: Union[str, SyntaxTheme]) -> SyntaxTheme:
        """Get a syntax theme instance."""
        if isinstance(name, SyntaxTheme):
            return name
        theme: SyntaxTheme
        if name in RICH_SYNTAX_THEMES:
            theme = ANSISyntaxTheme(RICH_SYNTAX_THEMES[name])
        else:
            theme = PygmentsSyntaxTheme(name)
        return theme

    def __init__(
        self,
        code: str,
        lexer: Union[Lexer, str],
        *,
        theme: Union[str, SyntaxTheme] = DEFAULT_THEME,
        dedent: bool = False,
        line_numbers: bool = False,
        start_line: int = 1,
        line_range: Optional[Tuple[Optional[int], Optional[int]]] = None,
        highlight_lines: Optional[Set[int]] = None,
        code_width: Optional[int] = None,
        tab_size: int = 4,
        word_wrap: bool = False,
        background_color: Optional[str] = None,
        indent_guides: bool = False,
        padding: PaddingDimensions = 0,
    ) -> None:
        self.code = code
        self._lexer = lexer
        self.dedent = dedent
        self.line_numbers = line_numbers
        self.start_line = start_line
        self.line_range = line_range
        self.highlight_lines = highlight_lines or set()
        self.code_width = code_width
        self.tab_size = tab_size
        self.word_wrap = word_wrap
        self.background_color = background_color
        self.background_style = (
            Style(bgcolor=background_color) if background_color else Style()
        )
        self.indent_guides = indent_guides
        self.padding = padding

        self._theme = self.get_theme(theme)
        self._stylized_ranges: List[_SyntaxHighlightRange] = []

    @classmethod
    def from_path(
        cls,
        path: str,
        encoding: str = "utf-8",
        lexer: Optional[Union[Lexer, str]] = None,
        theme: Union[str, SyntaxTheme] = DEFAULT_THEME,
        dedent: bool = False,
        line_numbers: bool = False,
        line_range: Optional[Tuple[int, int]] = None,
        start_line: int = 1,
        highlight_lines: Optional[Set[int]] = None,
        code_width: Optional[int] = None,
        tab_size: int = 4,
        word_wrap: bool = False,
        background_color: Optional[str] = None,
        indent_guides: bool = False,
        padding: PaddingDimensions = 0,
    ) -> "Syntax":
        """Construct a Syntax object from a file.

        Args:
            path (str): Path to file to highlight.
            encoding (str): Encoding of file.
            lexer (str | Lexer, optional): Lexer to use. If None, lexer will be auto-detected from path/file content.
            theme (str, optional): Color theme, aka Pygments style (see https://pygments.org/docs/styles/#getting-a-list-of-available-styles). Defaults to "emacs".
            dedent (bool, optional): Enable stripping of initial whitespace. Defaults to True.
            line_numbers (bool, optional): Enable rendering of line numbers. Defaults to False.
            start_line (int, optional): Starting number for line numbers. Defaults to 1.
            line_range (Tuple[int, int], optional): If given should be a tuple of the start and end line to render.
            highlight_lines (Set[int]): A set of line numbers to highlight.
            code_width: Width of code to render (not including line numbers), or ``None`` to use all available width.
            tab_size (int, optional): Size of tabs. Defaults to 4.
            word_wrap (bool, optional): Enable word wrapping of code.
            background_color (str, optional): Optional background color, or None to use theme color. Defaults to None.
            indent_guides (bool, optional): Show indent guides. Defaults to False.
            padding (PaddingDimensions): Padding to apply around the syntax. Defaults to 0 (no padding).

        Returns:
            [Syntax]: A Syntax object that may be printed to the console
        """
        code = Path(path).read_text(encoding=encoding)

        if not lexer:
            lexer = cls.guess_lexer(path, code=code)

        return cls(
            code,
            lexer,
            theme=theme,
            dedent=dedent,
            line_numbers=line_numbers,
            line_range=line_range,
            start_line=start_line,
            highlight_lines=highlight_lines,
            code_width=code_width,
            tab_size=tab_size,
            word_wrap=word_wrap,
            background_color=background_color,
            indent_guides=indent_guides,
            padding=padding,
        )

    @classmethod
    def guess_lexer(cls, path: str, code: Optional[str] = None) -> str:
        """Guess the alias of the Pygments lexer to use based on a path and an optional string of code.
        If code is supplied, it will use a combination of the code and the filename to determine the
        best lexer to use. For example, if the file is ``index.html`` and the file contains Django
        templating syntax, then "html+django" will be returned. If the file is ``index.html``, and no
        templating language is used, the "html" lexer will be used. If no string of code
        is supplied, the lexer will be chosen based on the file extension..

        Args:
             path (AnyStr): The path to the file containing the code you wish to know the lexer for.
             code (str, optional): Optional string of code that will be used as a fallback if no lexer
                is found for the supplied path.

        Returns:
            str: The name of the Pygments lexer that best matches the supplied path/code.
        """
        lexer: Optional[Lexer] = None
        lexer_name = "default"
        if code:
            try:
                lexer = guess_lexer_for_filename(path, code)
            except ClassNotFound:
                pass

        if not lexer:
            try:
                _, ext = os.path.splitext(path)
                if ext:
                    extension = ext.lstrip(".").lower()
                    lexer = get_lexer_by_name(extension)
            except ClassNotFound:
                pass

        if lexer:
            if lexer.aliases:
                lexer_name = lexer.aliases[0]
            else:
                lexer_name = lexer.name

        return lexer_name

    def _get_base_style(self) -> Style:
        """Get the base style."""
        default_style = self._theme.get_background_style() + self.background_style
        return default_style

    def _get_token_color(self, token_type: TokenType) -> Optional[Color]:
        """Get a color (if any) for the given token.

        Args:
            token_type (TokenType): A token type tuple from Pygments.

        Returns:
            Optional[Color]: Color from theme, or None for no color.
        """
        style = self._theme.get_style_for_token(token_type)
        return style.color

    @property
    def lexer(self) -> Optional[Lexer]:
        """The lexer for this syntax, or None if no lexer was found.

        Tries to find the lexer by name if a string was passed to the constructor.
        """

        if isinstance(self._lexer, Lexer):
            return self._lexer
        try:
            return get_lexer_by_name(
                self._lexer,
                stripnl=False,
                ensurenl=True,
                tabsize=self.tab_size,
            )
        except ClassNotFound:
            return None

    @property
    def default_lexer(self) -> Lexer:
        """A Pygments Lexer to use if one is not specified or invalid."""
        return get_lexer_by_name(
            "text",
            stripnl=False,
            ensurenl=True,
            tabsize=self.tab_size,
        )

    def highlight(
        self,
        code: str,
        line_range: Optional[Tuple[Optional[int], Optional[int]]] = None,
    ) -> Text:
        """Highlight code and return a Text instance.

        Args:
            code (str): Code to highlight.
            line_range(Tuple[int, int], optional): Optional line range to highlight.

        Returns:
            Text: A text instance containing highlighted syntax.
        """

        base_style = self._get_base_style()
        justify: JustifyMethod = (
            "default" if base_style.transparent_background else "left"
        )

        text = Text(
            justify=justify,
            style=base_style,
            tab_size=self.tab_size,
            no_wrap=not self.word_wrap,
        )
        _get_theme_style = self._theme.get_style_for_token

        lexer = self.lexer or self.default_lexer

        if lexer is None:
            text.append(code)
        else:
            if line_range:
                # More complicated path to only stylize a portion of the code
                # This speeds up further operations as there are less spans to process
                line_start, line_end = line_range

                def line_tokenize() -> Iterable[Tuple[Any, str]]:
                    """Split tokens to one per line."""
                    assert lexer  # required to make MyPy happy - we know lexer is not None at this point

                    for token_type, token in lexer.get_tokens(code):
                        while token:
                            line_token, new_line, token = token.partition("\n")
                            yield token_type, line_token + new_line

                def tokens_to_spans() -> Iterable[Tuple[str, Optional[Style]]]:
                    """Convert tokens to spans."""
                    tokens = iter(line_tokenize())
                    line_no = 0
                    _line_start = line_start - 1 if line_start else 0

                    # Skip over tokens until line start
                    while line_no < _line_start:
                        try:
                            _token_type, token = next(tokens)
                        except StopIteration:
                            break
                        yield (token, None)
                        if token.endswith("\n"):
                            line_no += 1
                    # Generate spans until line end
                    for token_type, token in tokens:
                        yield (token, _get_theme_style(token_type))
                        if token.endswith("\n"):
                            line_no += 1
                            if line_end and line_no >= line_end:
                                break

                text.append_tokens(tokens_to_spans())

            else:
                text.append_tokens(
                    (token, _get_theme_style(token_type))
                    for token_type, token in lexer.get_tokens(code)
                )
            if self.background_color is not None:
                text.stylize(f"on {self.background_color}")

        if self._stylized_ranges:
            self._apply_stylized_ranges(text)

        return text

    def stylize_range(
        self,
        style: StyleType,
        start: SyntaxPosition,
        end: SyntaxPosition,
        style_before: bool = False,
    ) -> None:
        """
        Adds a custom style on a part of the code, that will be applied to the syntax display when it's rendered.
        Line numbers are 1-based, while column indexes are 0-based.

        Args:
            style (StyleType): The style to apply.
            start (Tuple[int, int]): The start of the range, in the form `[line number, column index]`.
            end (Tuple[int, int]): The end of the range, in the form `[line number, column index]`.
            style_before (bool): Apply the style before any existing styles.
        """
        self._stylized_ranges.append(
            _SyntaxHighlightRange(style, start, end, style_before)
        )

    def _get_line_numbers_color(self, blend: float = 0.3) -> Color:
        background_style = self._theme.get_background_style() + self.background_style
        background_color = background_style.bgcolor
        if background_color is None or background_color.is_system_defined:
            return Color.default()
        foreground_color = self._get_token_color(Token.Text)
        if foreground_color is None or foreground_color.is_system_defined:
            return foreground_color or Color.default()
        new_color = blend_rgb(
            background_color.get_truecolor(),
            foreground_color.get_truecolor(),
            cross_fade=blend,
        )
        return Color.from_triplet(new_color)

    @property
    def _numbers_column_width(self) -> int:
        """Get the number of characters used to render the numbers column."""
        column_width = 0
        if self.line_numbers:
            column_width = (
                len(str(self.start_line + self.code.count("\n")))
                + NUMBERS_COLUMN_DEFAULT_PADDING
            )
        return column_width

    def _get_number_styles(self, console: Console) -> Tuple[Style, Style, Style]:
        """Get background, number, and highlight styles for line numbers."""
        background_style = self._get_base_style()
        if background_style.transparent_background:
            return Style.null(), Style(dim=True), Style.null()
        if console.color_system in ("256", "truecolor"):
            number_style = Style.chain(
                background_style,
                self._theme.get_style_for_token(Token.Text),
                Style(color=self._get_line_numbers_color()),
                self.background_style,
            )
            highlight_number_style = Style.chain(
                background_style,
                self._theme.get_style_for_token(Token.Text),
                Style(bold=True, color=self._get_line_numbers_color(0.9)),
                self.background_style,
            )
        else:
            number_style = background_style + Style(dim=True)
            highlight_number_style = background_style + Style(dim=False)
        return background_style, number_style, highlight_number_style

    def __rich_measure__(
        self, console: "Console", options: "ConsoleOptions"
    ) -> "Measurement":
        _, right, _, left = Padding.unpack(self.padding)
        padding = left + right
        if self.code_width is not None:
            width = self.code_width + self._numbers_column_width + padding + 1
            return Measurement(self._numbers_column_width, width)
        lines = self.code.splitlines()
        width = (
            self._numbers_column_width
            + padding
            + (max(cell_len(line) for line in lines) if lines else 0)
        )
        if self.line_numbers:
            width += 1
        return Measurement(self._numbers_column_width, width)

    def __rich_console__(
        self, console: Console, options: ConsoleOptions
    ) -> RenderResult:
        segments = Segments(self._get_syntax(console, options))
        if self.padding:
            yield Padding(segments, style=self._get_base_style(), pad=self.padding)
        else:
            yield segments

    def _get_syntax(
        self,
        console: Console,
        options: ConsoleOptions,
    ) -> Iterable[Segment]:
        """
        Get the Segments for the Syntax object, excluding any vertical/horizontal padding
        """
        transparent_background = self._get_base_style().transparent_background
        code_width = (
            (
                (options.max_width - self._numbers_column_width - 1)
                if self.line_numbers
                else options.max_width
            )
            if self.code_width is None
            else self.code_width
        )

        ends_on_nl, processed_code = self._process_code(self.code)
        text = self.highlight(processed_code, self.line_range)

        if not self.line_numbers and not self.word_wrap and not self.line_range:
            if not ends_on_nl:
                text.remove_suffix("\n")
            # Simple case of just rendering text
            style = (
                self._get_base_style()
                + self._theme.get_style_for_token(Comment)
                + Style(dim=True)
                + self.background_style
            )
            if self.indent_guides and not options.ascii_only:
                text = text.with_indent_guides(self.tab_size, style=style)
                text.overflow = "crop"
            if style.transparent_background:
                yield from console.render(
                    text, options=options.update(width=code_width)
                )
            else:
                syntax_lines = console.render_lines(
                    text,
                    options.update(width=code_width, height=None, justify="left"),
                    style=self.background_style,
                    pad=True,
                    new_lines=True,
                )
                for syntax_line in syntax_lines:
                    yield from syntax_line
            return

        start_line, end_line = self.line_range or (None, None)
        line_offset = 0
        if start_line:
            line_offset = max(0, start_line - 1)
        lines: Union[List[Text], Lines] = text.split("\n", allow_blank=ends_on_nl)
        if self.line_range:
            if line_offset > len(lines):
                return
            lines = lines[line_offset:end_line]

        if self.indent_guides and not options.ascii_only:
            style = (
                self._get_base_style()
                + self._theme.get_style_for_token(Comment)
                + Style(dim=True)
                + self.background_style
            )
            lines = (
                Text("\n")
                .join(lines)
                .with_indent_guides(self.tab_size, style=style + Style(italic=False))
                .split("\n", allow_blank=True)
            )

        numbers_column_width = self._numbers_column_width
        render_options = options.update(width=code_width)

        highlight_line = self.highlight_lines.__contains__
        _Segment = Segment
        new_line = _Segment("\n")

        line_pointer = "> " if options.legacy_windows else "❱ "

        (
            background_style,
            number_style,
            highlight_number_style,
        ) = self._get_number_styles(console)

        for line_no, line in enumerate(lines, self.start_line + line_offset):
            if self.word_wrap:
                wrapped_lines = console.render_lines(
                    line,
                    render_options.update(height=None, justify="left"),
                    style=background_style,
                    pad=not transparent_background,
                )
            else:
                segments = list(line.render(console, end=""))
                if options.no_wrap:
                    wrapped_lines = [segments]
                else:
                    wrapped_lines = [
                        _Segment.adjust_line_length(
                            segments,
                            render_options.max_width,
                            style=background_style,
                            pad=not transparent_background,
                        )
                    ]

            if self.line_numbers:
                wrapped_line_left_pad = _Segment(
                    " " * numbers_column_width + " ", background_style
                )
                for first, wrapped_line in loop_first(wrapped_lines):
                    if first:
                        line_column = str(line_no).rjust(numbers_column_width - 2) + " "
                        if highlight_line(line_no):
                            yield _Segment(line_pointer, Style(color="red"))
                            yield _Segment(line_column, highlight_number_style)
                        else:
                            yield _Segment("  ", highlight_number_style)
                            yield _Segment(line_column, number_style)
                    else:
                        yield wrapped_line_left_pad
                    yield from wrapped_line
                    yield new_line
            else:
                for wrapped_line in wrapped_lines:
                    yield from wrapped_line
                    yield new_line

    def _apply_stylized_ranges(self, text: Text) -> None:
        """
        Apply stylized ranges to a text instance,
        using the given code to determine the right portion to apply the style to.

        Args:
            text (Text): Text instance to apply the style to.
        """
        code = text.plain
        newlines_offsets = [
            # Let's add outer boundaries at each side of the list:
            0,
            # N.B. using "\n" here is much faster than using metacharacters such as "^" or "\Z":
            *[
                match.start() + 1
                for match in re.finditer("\n", code, flags=re.MULTILINE)
            ],
            len(code) + 1,
        ]

        for stylized_range in self._stylized_ranges:
            start = _get_code_index_for_syntax_position(
                newlines_offsets, stylized_range.start
            )
            end = _get_code_index_for_syntax_position(
                newlines_offsets, stylized_range.end
            )
            if start is not None and end is not None:
                if stylized_range.style_before:
                    text.stylize_before(stylized_range.style, start, end)
                else:
                    text.stylize(stylized_range.style, start, end)

    def _process_code(self, code: str) -> Tuple[bool, str]:
        """
        Applies various processing to a raw code string
        (normalises it so it always ends with a line return, dedents it if necessary, etc.)

        Args:
            code (str): The raw code string to process

        Returns:
            Tuple[bool, str]: the boolean indicates whether the raw code ends with a line return,
                while the string is the processed code.
        """
        ends_on_nl = code.endswith("\n")
        processed_code = code if ends_on_nl else code + "\n"
        processed_code = (
            textwrap.dedent(processed_code) if self.dedent else processed_code
        )
        processed_code = processed_code.expandtabs(self.tab_size)
        return ends_on_nl, processed_code


def _get_code_index_for_syntax_position(
    newlines_offsets: Sequence[int], position: SyntaxPosition
) -> Optional[int]:
    """
    Returns the index of the code string for the given positions.

    Args:
        newlines_offsets (Sequence[int]): The offset of each newline character found in the code snippet.
        position (SyntaxPosition): The position to search for.

    Returns:
        Optional[int]: The index of the code string for this position, or `None`
            if the given position's line number is out of range (if it's the column that is out of range
            we silently clamp its value so that it reaches the end of the line)
    """
    lines_count = len(newlines_offsets)

    line_number, column_index = position
    if line_number > lines_count or len(newlines_offsets) < (line_number + 1):
        return None  # `line_number` is out of range
    line_index = line_number - 1
    line_length = newlines_offsets[line_index + 1] - newlines_offsets[line_index] - 1
    # If `column_index` is out of range: let's silently clamp it:
    column_index = min(line_length, column_index)
    return newlines_offsets[line_index] + column_index


if __name__ == "__main__":  # pragma: no cover
    import argparse
    import sys

    parser = argparse.ArgumentParser(
        description="Render syntax to the console with Rich"
    )
    parser.add_argument(
        "path",
        metavar="PATH",
        help="path to file, or - for stdin",
    )
    parser.add_argument(
        "-c",
        "--force-color",
        dest="force_color",
        action="store_true",
        default=None,
        help="force color for non-terminals",
    )
    parser.add_argument(
        "-i",
        "--indent-guides",
        dest="indent_guides",
        action="store_true",
        default=False,
        help="display indent guides",
    )
    parser.add_argument(
        "-l",
        "--line-numbers",
        dest="line_numbers",
        action="store_true",
        help="render line numbers",
    )
    parser.add_argument(
        "-w",
        "--width",
        type=int,
        dest="width",
        default=None,
        help="width of output (default will auto-detect)",
    )
    parser.add_argument(
        "-r",
        "--wrap",
        dest="word_wrap",
        action="store_true",
        default=False,
        help="word wrap long lines",
    )
    parser.add_argument(
        "-s",
        "--soft-wrap",
        action="store_true",
        dest="soft_wrap",
        default=False,
        help="enable soft wrapping mode",
    )
    parser.add_argument(
        "-t", "--theme", dest="theme", default="monokai", help="pygments theme"
    )
    parser.add_argument(
        "-b",
        "--background-color",
        dest="background_color",
        default=None,
        help="Override background color",
    )
    parser.add_argument(
        "-x",
        "--lexer",
        default=None,
        dest="lexer_name",
        help="Lexer name",
    )
    parser.add_argument(
        "-p", "--padding", type=int, default=0, dest="padding", help="Padding"
    )
    parser.add_argument(
        "--highlight-line",
        type=int,
        default=None,
        dest="highlight_line",
        help="The line number (not index!) to highlight",
    )
    args = parser.parse_args()

    from pip._vendor.rich.console import Console

    console = Console(force_terminal=args.force_color, width=args.width)

    if args.path == "-":
        code = sys.stdin.read()
        syntax = Syntax(
            code=code,
            lexer=args.lexer_name,
            line_numbers=args.line_numbers,
            word_wrap=args.word_wrap,
            theme=args.theme,
            background_color=args.background_color,
            indent_guides=args.indent_guides,
            padding=args.padding,
            highlight_lines={args.highlight_line},
        )
    else:
        syntax = Syntax.from_path(
            args.path,
            lexer=args.lexer_name,
            line_numbers=args.line_numbers,
            word_wrap=args.word_wrap,
            theme=args.theme,
            background_color=args.background_color,
            indent_guides=args.indent_guides,
            padding=args.padding,
            highlight_lines={args.highlight_line},
        )
    console.print(syntax, soft_wrap=args.soft_wrap)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\solvers\pde.py ===
"""
This module contains pdsolve() and different helper functions that it
uses. It is heavily inspired by the ode module and hence the basic
infrastructure remains the same.

**Functions in this module**

    These are the user functions in this module:

    - pdsolve()     - Solves PDE's
    - classify_pde() - Classifies PDEs into possible hints for dsolve().
    - pde_separate() - Separate variables in partial differential equation either by
                       additive or multiplicative separation approach.

    These are the helper functions in this module:

    - pde_separate_add() - Helper function for searching additive separable solutions.
    - pde_separate_mul() - Helper function for searching multiplicative
                           separable solutions.

**Currently implemented solver methods**

The following methods are implemented for solving partial differential
equations.  See the docstrings of the various pde_hint() functions for
more information on each (run help(pde)):

  - 1st order linear homogeneous partial differential equations
    with constant coefficients.
  - 1st order linear general partial differential equations
    with constant coefficients.
  - 1st order linear partial differential equations with
    variable coefficients.

"""
from functools import reduce

from itertools import combinations_with_replacement
from sympy.simplify import simplify  # type: ignore
from sympy.core import Add, S
from sympy.core.function import Function, expand, AppliedUndef, Subs
from sympy.core.relational import Equality, Eq
from sympy.core.symbol import Symbol, Wild, symbols
from sympy.functions import exp
from sympy.integrals.integrals import Integral, integrate
from sympy.utilities.iterables import has_dups, is_sequence
from sympy.utilities.misc import filldedent

from sympy.solvers.deutils import _preprocess, ode_order, _desolve
from sympy.solvers.solvers import solve
from sympy.simplify.radsimp import collect

import operator


allhints = (
    "1st_linear_constant_coeff_homogeneous",
    "1st_linear_constant_coeff",
    "1st_linear_constant_coeff_Integral",
    "1st_linear_variable_coeff"
    )


def pdsolve(eq, func=None, hint='default', dict=False, solvefun=None, **kwargs):
    """
    Solves any (supported) kind of partial differential equation.

    **Usage**

        pdsolve(eq, f(x,y), hint) -> Solve partial differential equation
        eq for function f(x,y), using method hint.

    **Details**

        ``eq`` can be any supported partial differential equation (see
            the pde docstring for supported methods).  This can either
            be an Equality, or an expression, which is assumed to be
            equal to 0.

        ``f(x,y)`` is a function of two variables whose derivatives in that
            variable make up the partial differential equation. In many
            cases it is not necessary to provide this; it will be autodetected
            (and an error raised if it could not be detected).

        ``hint`` is the solving method that you want pdsolve to use.  Use
            classify_pde(eq, f(x,y)) to get all of the possible hints for
            a PDE.  The default hint, 'default', will use whatever hint
            is returned first by classify_pde().  See Hints below for
            more options that you can use for hint.

        ``solvefun`` is the convention used for arbitrary functions returned
            by the PDE solver. If not set by the user, it is set by default
            to be F.

    **Hints**

        Aside from the various solving methods, there are also some
        meta-hints that you can pass to pdsolve():

        "default":
                This uses whatever hint is returned first by
                classify_pde(). This is the default argument to
                pdsolve().

        "all":
                To make pdsolve apply all relevant classification hints,
                use pdsolve(PDE, func, hint="all").  This will return a
                dictionary of hint:solution terms.  If a hint causes
                pdsolve to raise the NotImplementedError, value of that
                hint's key will be the exception object raised.  The
                dictionary will also include some special keys:

                - order: The order of the PDE.  See also ode_order() in
                  deutils.py
                - default: The solution that would be returned by
                  default.  This is the one produced by the hint that
                  appears first in the tuple returned by classify_pde().

        "all_Integral":
                This is the same as "all", except if a hint also has a
                corresponding "_Integral" hint, it only returns the
                "_Integral" hint.  This is useful if "all" causes
                pdsolve() to hang because of a difficult or impossible
                integral.  This meta-hint will also be much faster than
                "all", because integrate() is an expensive routine.

        See also the classify_pde() docstring for more info on hints,
        and the pde docstring for a list of all supported hints.

    **Tips**
        - You can declare the derivative of an unknown function this way:

            >>> from sympy import Function, Derivative
            >>> from sympy.abc import x, y # x and y are the independent variables
            >>> f = Function("f")(x, y) # f is a function of x and y
            >>> # fx will be the partial derivative of f with respect to x
            >>> fx = Derivative(f, x)
            >>> # fy will be the partial derivative of f with respect to y
            >>> fy = Derivative(f, y)

        - See test_pde.py for many tests, which serves also as a set of
          examples for how to use pdsolve().
        - pdsolve always returns an Equality class (except for the case
          when the hint is "all" or "all_Integral"). Note that it is not possible
          to get an explicit solution for f(x, y) as in the case of ODE's
        - Do help(pde.pde_hintname) to get help more information on a
          specific hint


    Examples
    ========

    >>> from sympy.solvers.pde import pdsolve
    >>> from sympy import Function, Eq
    >>> from sympy.abc import x, y
    >>> f = Function('f')
    >>> u = f(x, y)
    >>> ux = u.diff(x)
    >>> uy = u.diff(y)
    >>> eq = Eq(1 + (2*(ux/u)) + (3*(uy/u)), 0)
    >>> pdsolve(eq)
    Eq(f(x, y), F(3*x - 2*y)*exp(-2*x/13 - 3*y/13))

    """

    if not solvefun:
        solvefun = Function('F')

    # See the docstring of _desolve for more details.
    hints = _desolve(eq, func=func, hint=hint, simplify=True,
                     type='pde', **kwargs)
    eq = hints.pop('eq', False)
    all_ = hints.pop('all', False)

    if all_:
        # TODO : 'best' hint should be implemented when adequate
        # number of hints are added.
        pdedict = {}
        failed_hints = {}
        gethints = classify_pde(eq, dict=True)
        pdedict.update({'order': gethints['order'],
                        'default': gethints['default']})
        for hint in hints:
            try:
                rv = _helper_simplify(eq, hint, hints[hint]['func'],
                    hints[hint]['order'], hints[hint][hint], solvefun)
            except NotImplementedError as detail:
                failed_hints[hint] = detail
            else:
                pdedict[hint] = rv
        pdedict.update(failed_hints)
        return pdedict

    else:
        return _helper_simplify(eq, hints['hint'], hints['func'],
                                hints['order'], hints[hints['hint']], solvefun)


def _helper_simplify(eq, hint, func, order, match, solvefun):
    """Helper function of pdsolve that calls the respective
    pde functions to solve for the partial differential
    equations. This minimizes the computation in
    calling _desolve multiple times.
    """
    solvefunc = globals()["pde_" + hint.removesuffix("_Integral")]
    return _handle_Integral(solvefunc(eq, func, order,
        match, solvefun), func, order, hint)


def _handle_Integral(expr, func, order, hint):
    r"""
    Converts a solution with integrals in it into an actual solution.

    Simplifies the integral mainly using doit()
    """
    if hint.endswith("_Integral"):
        return expr

    elif hint == "1st_linear_constant_coeff":
        return simplify(expr.doit())

    else:
        return expr


def classify_pde(eq, func=None, dict=False, *, prep=True, **kwargs):
    """
    Returns a tuple of possible pdsolve() classifications for a PDE.

    The tuple is ordered so that first item is the classification that
    pdsolve() uses to solve the PDE by default.  In general,
    classifications near the beginning of the list will produce
    better solutions faster than those near the end, though there are
    always exceptions.  To make pdsolve use a different classification,
    use pdsolve(PDE, func, hint=<classification>).  See also the pdsolve()
    docstring for different meta-hints you can use.

    If ``dict`` is true, classify_pde() will return a dictionary of
    hint:match expression terms. This is intended for internal use by
    pdsolve().  Note that because dictionaries are ordered arbitrarily,
    this will most likely not be in the same order as the tuple.

    You can get help on different hints by doing help(pde.pde_hintname),
    where hintname is the name of the hint without "_Integral".

    See sympy.pde.allhints or the sympy.pde docstring for a list of all
    supported hints that can be returned from classify_pde.


    Examples
    ========

    >>> from sympy.solvers.pde import classify_pde
    >>> from sympy import Function, Eq
    >>> from sympy.abc import x, y
    >>> f = Function('f')
    >>> u = f(x, y)
    >>> ux = u.diff(x)
    >>> uy = u.diff(y)
    >>> eq = Eq(1 + (2*(ux/u)) + (3*(uy/u)), 0)
    >>> classify_pde(eq)
    ('1st_linear_constant_coeff_homogeneous',)
    """

    if func and len(func.args) != 2:
        raise NotImplementedError("Right now only partial "
            "differential equations of two variables are supported")

    if prep or func is None:
        prep, func_ = _preprocess(eq, func)
        if func is None:
            func = func_

    if isinstance(eq, Equality):
        if eq.rhs != 0:
            return classify_pde(eq.lhs - eq.rhs, func)
        eq = eq.lhs

    f = func.func
    x = func.args[0]
    y = func.args[1]
    fx = f(x,y).diff(x)
    fy = f(x,y).diff(y)

    # TODO : For now pde.py uses support offered by the ode_order function
    # to find the order with respect to a multi-variable function. An
    # improvement could be to classify the order of the PDE on the basis of
    # individual variables.
    order = ode_order(eq, f(x,y))

    # hint:matchdict or hint:(tuple of matchdicts)
    # Also will contain "default":<default hint> and "order":order items.
    matching_hints = {'order': order}

    if not order:
        if dict:
            matching_hints["default"] = None
            return matching_hints
        return ()

    eq = expand(eq)

    a = Wild('a', exclude = [f(x,y)])
    b = Wild('b', exclude = [f(x,y), fx, fy, x, y])
    c = Wild('c', exclude = [f(x,y), fx, fy, x, y])
    d = Wild('d', exclude = [f(x,y), fx, fy, x, y])
    e = Wild('e', exclude = [f(x,y), fx, fy])
    n = Wild('n', exclude = [x, y])
    # Try removing the smallest power of f(x,y)
    # from the highest partial derivatives of f(x,y)
    reduced_eq = eq
    if eq.is_Add:
        power = None
        for i in set(combinations_with_replacement((x,y), order)):
            coeff = eq.coeff(f(x,y).diff(*i))
            if coeff == 1:
                continue
            match = coeff.match(a*f(x,y)**n)
            if match and match[a]:
                if power is None or match[n] < power:
                    power = match[n]
        if power:
            den = f(x,y)**power
            reduced_eq = Add(*[arg/den for arg in eq.args])

    if order == 1:
        reduced_eq = collect(reduced_eq, f(x, y))
        r = reduced_eq.match(b*fx + c*fy + d*f(x,y) + e)
        if r:
            if not r[e]:
                ## Linear first-order homogeneous partial-differential
                ## equation with constant coefficients
                r.update({'b': b, 'c': c, 'd': d})
                matching_hints["1st_linear_constant_coeff_homogeneous"] = r
            elif r[b]**2 + r[c]**2 != 0:
                ## Linear first-order general partial-differential
                ## equation with constant coefficients
                r.update({'b': b, 'c': c, 'd': d, 'e': e})
                matching_hints["1st_linear_constant_coeff"] = r
                matching_hints["1st_linear_constant_coeff_Integral"] = r

        else:
            b = Wild('b', exclude=[f(x, y), fx, fy])
            c = Wild('c', exclude=[f(x, y), fx, fy])
            d = Wild('d', exclude=[f(x, y), fx, fy])
            r = reduced_eq.match(b*fx + c*fy + d*f(x,y) + e)
            if r:
                r.update({'b': b, 'c': c, 'd': d, 'e': e})
                matching_hints["1st_linear_variable_coeff"] = r

    # Order keys based on allhints.
    rettuple = tuple(i for i in allhints if i in matching_hints)

    if dict:
        # Dictionaries are ordered arbitrarily, so make note of which
        # hint would come first for pdsolve().  Use an ordered dict in Py 3.
        matching_hints["default"] = None
        matching_hints["ordered_hints"] = rettuple
        for i in allhints:
            if i in matching_hints:
                matching_hints["default"] = i
                break
        return matching_hints
    return rettuple


def checkpdesol(pde, sol, func=None, solve_for_func=True):
    """
    Checks if the given solution satisfies the partial differential
    equation.

    pde is the partial differential equation which can be given in the
    form of an equation or an expression. sol is the solution for which
    the pde is to be checked. This can also be given in an equation or
    an expression form. If the function is not provided, the helper
    function _preprocess from deutils is used to identify the function.

    If a sequence of solutions is passed, the same sort of container will be
    used to return the result for each solution.

    The following methods are currently being implemented to check if the
    solution satisfies the PDE:

        1. Directly substitute the solution in the PDE and check. If the
           solution has not been solved for f, then it will solve for f
           provided solve_for_func has not been set to False.

    If the solution satisfies the PDE, then a tuple (True, 0) is returned.
    Otherwise a tuple (False, expr) where expr is the value obtained
    after substituting the solution in the PDE. However if a known solution
    returns False, it may be due to the inability of doit() to simplify it to zero.

    Examples
    ========

    >>> from sympy import Function, symbols
    >>> from sympy.solvers.pde import checkpdesol, pdsolve
    >>> x, y = symbols('x y')
    >>> f = Function('f')
    >>> eq = 2*f(x,y) + 3*f(x,y).diff(x) + 4*f(x,y).diff(y)
    >>> sol = pdsolve(eq)
    >>> assert checkpdesol(eq, sol)[0]
    >>> eq = x*f(x,y) + f(x,y).diff(x)
    >>> checkpdesol(eq, sol)
    (False, (x*F(4*x - 3*y) - 6*F(4*x - 3*y)/25 + 4*Subs(Derivative(F(_xi_1), _xi_1), _xi_1, 4*x - 3*y))*exp(-6*x/25 - 8*y/25))
    """

    # Converting the pde into an equation
    if not isinstance(pde, Equality):
        pde = Eq(pde, 0)

    # If no function is given, try finding the function present.
    if func is None:
        try:
            _, func = _preprocess(pde.lhs)
        except ValueError:
            funcs = [s.atoms(AppliedUndef) for s in (
                sol if is_sequence(sol, set) else [sol])]
            funcs = set().union(funcs)
            if len(funcs) != 1:
                raise ValueError(
                    'must pass func arg to checkpdesol for this case.')
            func = funcs.pop()

    # If the given solution is in the form of a list or a set
    # then return a list or set of tuples.
    if is_sequence(sol, set):
        return type(sol)([checkpdesol(
            pde, i, func=func,
            solve_for_func=solve_for_func) for i in sol])

    # Convert solution into an equation
    if not isinstance(sol, Equality):
        sol = Eq(func, sol)
    elif sol.rhs == func:
        sol = sol.reversed

    # Try solving for the function
    solved = sol.lhs == func and not sol.rhs.has(func)
    if solve_for_func and not solved:
        solved = solve(sol, func)
        if solved:
            if len(solved) == 1:
                return checkpdesol(pde, Eq(func, solved[0]),
                    func=func, solve_for_func=False)
            else:
                return checkpdesol(pde, [Eq(func, t) for t in solved],
                    func=func, solve_for_func=False)

    # try direct substitution of the solution into the PDE and simplify
    if sol.lhs == func:
        pde = pde.lhs - pde.rhs
        s = simplify(pde.subs(func, sol.rhs).doit())
        return s is S.Zero, s

    raise NotImplementedError(filldedent('''
        Unable to test if %s is a solution to %s.''' % (sol, pde)))



def pde_1st_linear_constant_coeff_homogeneous(eq, func, order, match, solvefun):
    r"""
    Solves a first order linear homogeneous
    partial differential equation with constant coefficients.

    The general form of this partial differential equation is

    .. math:: a \frac{\partial f(x,y)}{\partial x}
              + b \frac{\partial f(x,y)}{\partial y} + c f(x,y) = 0

    where `a`, `b` and `c` are constants.

    The general solution is of the form:

    .. math::
        f(x, y) = F(- a y + b x ) e^{- \frac{c (a x + b y)}{a^2 + b^2}}

    and can be found in SymPy with ``pdsolve``::

        >>> from sympy.solvers import pdsolve
        >>> from sympy.abc import x, y, a, b, c
        >>> from sympy import Function, pprint
        >>> f = Function('f')
        >>> u = f(x,y)
        >>> ux = u.diff(x)
        >>> uy = u.diff(y)
        >>> genform = a*ux + b*uy + c*u
        >>> pprint(genform)
          d               d
        a*--(f(x, y)) + b*--(f(x, y)) + c*f(x, y)
          dx              dy

        >>> pprint(pdsolve(genform))
                                 -c*(a*x + b*y)
                                 ---------------
                                      2    2
                                     a  + b
        f(x, y) = F(-a*y + b*x)*e

    Examples
    ========

    >>> from sympy import pdsolve
    >>> from sympy import Function, pprint
    >>> from sympy.abc import x,y
    >>> f = Function('f')
    >>> pdsolve(f(x,y) + f(x,y).diff(x) + f(x,y).diff(y))
    Eq(f(x, y), F(x - y)*exp(-x/2 - y/2))
    >>> pprint(pdsolve(f(x,y) + f(x,y).diff(x) + f(x,y).diff(y)))
                          x   y
                        - - - -
                          2   2
    f(x, y) = F(x - y)*e

    References
    ==========

    - Viktor Grigoryan, "Partial Differential Equations"
      Math 124A - Fall 2010, pp.7

    """
    # TODO : For now homogeneous first order linear PDE's having
    # two variables are implemented. Once there is support for
    # solving systems of ODE's, this can be extended to n variables.

    f = func.func
    x = func.args[0]
    y = func.args[1]
    b = match[match['b']]
    c = match[match['c']]
    d = match[match['d']]
    return Eq(f(x,y), exp(-S(d)/(b**2 + c**2)*(b*x + c*y))*solvefun(c*x - b*y))


def pde_1st_linear_constant_coeff(eq, func, order, match, solvefun):
    r"""
    Solves a first order linear partial differential equation
    with constant coefficients.

    The general form of this partial differential equation is

    .. math:: a \frac{\partial f(x,y)}{\partial x}
              + b \frac{\partial f(x,y)}{\partial y}
              + c f(x,y) = G(x,y)

    where `a`, `b` and `c` are constants and `G(x, y)` can be an arbitrary
    function in `x` and `y`.

    The general solution of the PDE is:

    .. math::
        f(x, y) = \left. \left[F(\eta) + \frac{1}{a^2 + b^2}
        \int\limits^{a x + b y} G\left(\frac{a \xi + b \eta}{a^2 + b^2},
        \frac{- a \eta + b \xi}{a^2 + b^2} \right)
        e^{\frac{c \xi}{a^2 + b^2}}\, d\xi\right]
        e^{- \frac{c \xi}{a^2 + b^2}}
        \right|_{\substack{\eta=- a y + b x\\ \xi=a x + b y }}\, ,

    where `F(\eta)` is an arbitrary single-valued function. The solution
    can be found in SymPy with ``pdsolve``::

        >>> from sympy.solvers import pdsolve
        >>> from sympy.abc import x, y, a, b, c
        >>> from sympy import Function, pprint
        >>> f = Function('f')
        >>> G = Function('G')
        >>> u = f(x, y)
        >>> ux = u.diff(x)
        >>> uy = u.diff(y)
        >>> genform = a*ux + b*uy + c*u - G(x,y)
        >>> pprint(genform)
          d               d
        a*--(f(x, y)) + b*--(f(x, y)) + c*f(x, y) - G(x, y)
          dx              dy
        >>> pprint(pdsolve(genform, hint='1st_linear_constant_coeff_Integral'))
                  //          a*x + b*y                                             \         \|
                  ||              /                                                 |         ||
                  ||             |                                                  |         ||
                  ||             |                                      c*xi        |         ||
                  ||             |                                     -------      |         ||
                  ||             |                                      2    2      |         ||
                  ||             |      /a*xi + b*eta  -a*eta + b*xi\  a  + b       |         ||
                  ||             |     G|------------, -------------|*e        d(xi)|         ||
                  ||             |      |   2    2         2    2   |               |         ||
                  ||             |      \  a  + b         a  + b    /               |  -c*xi  ||
                  ||             |                                                  |  -------||
                  ||            /                                                   |   2    2||
                  ||                                                                |  a  + b ||
        f(x, y) = ||F(eta) + -------------------------------------------------------|*e       ||
                  ||                                  2    2                        |         ||
                  \\                                 a  + b                         /         /|eta=-a*y + b*x, xi=a*x + b*y

    Examples
    ========

    >>> from sympy.solvers.pde import pdsolve
    >>> from sympy import Function, pprint, exp
    >>> from sympy.abc import x,y
    >>> f = Function('f')
    >>> eq = -2*f(x,y).diff(x) + 4*f(x,y).diff(y) + 5*f(x,y) - exp(x + 3*y)
    >>> pdsolve(eq)
    Eq(f(x, y), (F(4*x + 2*y)*exp(x/2) + exp(x + 4*y)/15)*exp(-y))

    References
    ==========

    - Viktor Grigoryan, "Partial Differential Equations"
      Math 124A - Fall 2010, pp.7

    """

    # TODO : For now homogeneous first order linear PDE's having
    # two variables are implemented. Once there is support for
    # solving systems of ODE's, this can be extended to n variables.
    xi, eta = symbols("xi eta")
    f = func.func
    x = func.args[0]
    y = func.args[1]
    b = match[match['b']]
    c = match[match['c']]
    d = match[match['d']]
    e = -match[match['e']]
    expterm = exp(-S(d)/(b**2 + c**2)*xi)
    functerm = solvefun(eta)
    solvedict = solve((b*x + c*y - xi, c*x - b*y - eta), x, y)
    # Integral should remain as it is in terms of xi,
    # doit() should be done in _handle_Integral.
    genterm = (1/S(b**2 + c**2))*Integral(
        (1/expterm*e).subs(solvedict), (xi, b*x + c*y))
    return Eq(f(x,y), Subs(expterm*(functerm + genterm),
        (eta, xi), (c*x - b*y, b*x + c*y)))


def pde_1st_linear_variable_coeff(eq, func, order, match, solvefun):
    r"""
    Solves a first order linear partial differential equation
    with variable coefficients. The general form of this partial
    differential equation is

    .. math:: a(x, y) \frac{\partial f(x, y)}{\partial x}
                + b(x, y) \frac{\partial f(x, y)}{\partial y}
                + c(x, y) f(x, y) = G(x, y)

    where `a(x, y)`, `b(x, y)`, `c(x, y)` and `G(x, y)` are arbitrary
    functions in `x` and `y`. This PDE is converted into an ODE by
    making the following transformation:

    1. `\xi` as `x`

    2. `\eta` as the constant in the solution to the differential
       equation `\frac{dy}{dx} = -\frac{b}{a}`

    Making the previous substitutions reduces it to the linear ODE

    .. math:: a(\xi, \eta)\frac{du}{d\xi} + c(\xi, \eta)u - G(\xi, \eta) = 0

    which can be solved using ``dsolve``.

    >>> from sympy.abc import x, y
    >>> from sympy import Function, pprint
    >>> a, b, c, G, f= [Function(i) for i in ['a', 'b', 'c', 'G', 'f']]
    >>> u = f(x,y)
    >>> ux = u.diff(x)
    >>> uy = u.diff(y)
    >>> genform = a(x, y)*u + b(x, y)*ux + c(x, y)*uy - G(x,y)
    >>> pprint(genform)
                                         d                     d
    -G(x, y) + a(x, y)*f(x, y) + b(x, y)*--(f(x, y)) + c(x, y)*--(f(x, y))
                                         dx                    dy


    Examples
    ========

    >>> from sympy.solvers.pde import pdsolve
    >>> from sympy import Function, pprint
    >>> from sympy.abc import x,y
    >>> f = Function('f')
    >>> eq =  x*(u.diff(x)) - y*(u.diff(y)) + y**2*u - y**2
    >>> pdsolve(eq)
    Eq(f(x, y), F(x*y)*exp(y**2/2) + 1)

    References
    ==========

    - Viktor Grigoryan, "Partial Differential Equations"
      Math 124A - Fall 2010, pp.7

    """
    from sympy.solvers.ode import dsolve

    eta = symbols("eta")
    f = func.func
    x = func.args[0]
    y = func.args[1]
    b = match[match['b']]
    c = match[match['c']]
    d = match[match['d']]
    e = -match[match['e']]


    if not d:
         # To deal with cases like b*ux = e or c*uy = e
        if not (b and c):
            if c:
                try:
                    tsol = integrate(e/c, y)
                except NotImplementedError:
                    raise NotImplementedError("Unable to find a solution"
                        " due to inability of integrate")
                else:
                    return Eq(f(x,y), solvefun(x) + tsol)
            if b:
                try:
                    tsol = integrate(e/b, x)
                except NotImplementedError:
                    raise NotImplementedError("Unable to find a solution"
                        " due to inability of integrate")
                else:
                    return Eq(f(x,y), solvefun(y) + tsol)

    if not c:
        # To deal with cases when c is 0, a simpler method is used.
        # The PDE reduces to b*(u.diff(x)) + d*u = e, which is a linear ODE in x
        plode = f(x).diff(x)*b + d*f(x) - e
        sol = dsolve(plode, f(x))
        syms = sol.free_symbols - plode.free_symbols - {x, y}
        rhs = _simplify_variable_coeff(sol.rhs, syms, solvefun, y)
        return Eq(f(x, y), rhs)

    if not b:
        # To deal with cases when b is 0, a simpler method is used.
        # The PDE reduces to c*(u.diff(y)) + d*u = e, which is a linear ODE in y
        plode = f(y).diff(y)*c + d*f(y) - e
        sol = dsolve(plode, f(y))
        syms = sol.free_symbols - plode.free_symbols - {x, y}
        rhs = _simplify_variable_coeff(sol.rhs, syms, solvefun, x)
        return Eq(f(x, y), rhs)

    dummy = Function('d')
    h = (c/b).subs(y, dummy(x))
    sol = dsolve(dummy(x).diff(x) - h, dummy(x))
    if isinstance(sol, list):
        sol = sol[0]
    solsym = sol.free_symbols - h.free_symbols - {x, y}
    if len(solsym) == 1:
        solsym = solsym.pop()
        etat = (solve(sol, solsym)[0]).subs(dummy(x), y)
        ysub = solve(eta - etat, y)[0]
        deq = (b*(f(x).diff(x)) + d*f(x) - e).subs(y, ysub)
        final = (dsolve(deq, f(x), hint='1st_linear')).rhs
        if isinstance(final, list):
            final = final[0]
        finsyms = final.free_symbols - deq.free_symbols - {x, y}
        rhs = _simplify_variable_coeff(final, finsyms, solvefun, etat)
        return Eq(f(x, y), rhs)

    else:
        raise NotImplementedError("Cannot solve the partial differential equation due"
            " to inability of constantsimp")


def _simplify_variable_coeff(sol, syms, func, funcarg):
    r"""
    Helper function to replace constants by functions in 1st_linear_variable_coeff
    """
    eta = Symbol("eta")
    if len(syms) == 1:
        sym = syms.pop()
        final = sol.subs(sym, func(funcarg))

    else:
        for sym in syms:
            final = sol.subs(sym, func(funcarg))

    return simplify(final.subs(eta, funcarg))


def pde_separate(eq, fun, sep, strategy='mul'):
    """Separate variables in partial differential equation either by additive
    or multiplicative separation approach. It tries to rewrite an equation so
    that one of the specified variables occurs on a different side of the
    equation than the others.

    :param eq: Partial differential equation

    :param fun: Original function F(x, y, z)

    :param sep: List of separated functions [X(x), u(y, z)]

    :param strategy: Separation strategy. You can choose between additive
        separation ('add') and multiplicative separation ('mul') which is
        default.

    Examples
    ========

    >>> from sympy import E, Eq, Function, pde_separate, Derivative as D
    >>> from sympy.abc import x, t
    >>> u, X, T = map(Function, 'uXT')

    >>> eq = Eq(D(u(x, t), x), E**(u(x, t))*D(u(x, t), t))
    >>> pde_separate(eq, u(x, t), [X(x), T(t)], strategy='add')
    [exp(-X(x))*Derivative(X(x), x), exp(T(t))*Derivative(T(t), t)]

    >>> eq = Eq(D(u(x, t), x, 2), D(u(x, t), t, 2))
    >>> pde_separate(eq, u(x, t), [X(x), T(t)], strategy='mul')
    [Derivative(X(x), (x, 2))/X(x), Derivative(T(t), (t, 2))/T(t)]

    See Also
    ========
    pde_separate_add, pde_separate_mul
    """

    do_add = False
    if strategy == 'add':
        do_add = True
    elif strategy == 'mul':
        do_add = False
    else:
        raise ValueError('Unknown strategy: %s' % strategy)

    if isinstance(eq, Equality):
        if eq.rhs != 0:
            return pde_separate(Eq(eq.lhs - eq.rhs, 0), fun, sep, strategy)
    else:
        return pde_separate(Eq(eq, 0), fun, sep, strategy)

    if eq.rhs != 0:
        raise ValueError("Value should be 0")

    # Handle arguments
    orig_args = list(fun.args)
    subs_args = [arg for s in sep for arg in s.args]

    if do_add:
        functions = reduce(operator.add, sep)
    else:
        functions = reduce(operator.mul, sep)

    # Check whether variables match
    if len(subs_args) != len(orig_args):
        raise ValueError("Variable counts do not match")
    # Check for duplicate arguments like  [X(x), u(x, y)]
    if has_dups(subs_args):
        raise ValueError("Duplicate substitution arguments detected")
    # Check whether the variables match
    if set(orig_args) != set(subs_args):
        raise ValueError("Arguments do not match")

    # Substitute original function with separated...
    result = eq.lhs.subs(fun, functions).doit()

    # Divide by terms when doing multiplicative separation
    if not do_add:
        eq = 0
        for i in result.args:
            eq += i/functions
        result = eq

    svar = subs_args[0]
    dvar = subs_args[1:]
    return _separate(result, svar, dvar)


def pde_separate_add(eq, fun, sep):
    """
    Helper function for searching additive separable solutions.

    Consider an equation of two independent variables x, y and a dependent
    variable w, we look for the product of two functions depending on different
    arguments:

    `w(x, y, z) = X(x) + y(y, z)`

    Examples
    ========

    >>> from sympy import E, Eq, Function, pde_separate_add, Derivative as D
    >>> from sympy.abc import x, t
    >>> u, X, T = map(Function, 'uXT')

    >>> eq = Eq(D(u(x, t), x), E**(u(x, t))*D(u(x, t), t))
    >>> pde_separate_add(eq, u(x, t), [X(x), T(t)])
    [exp(-X(x))*Derivative(X(x), x), exp(T(t))*Derivative(T(t), t)]

    """
    return pde_separate(eq, fun, sep, strategy='add')


def pde_separate_mul(eq, fun, sep):
    """
    Helper function for searching multiplicative separable solutions.

    Consider an equation of two independent variables x, y and a dependent
    variable w, we look for the product of two functions depending on different
    arguments:

    `w(x, y, z) = X(x)*u(y, z)`

    Examples
    ========

    >>> from sympy import Function, Eq, pde_separate_mul, Derivative as D
    >>> from sympy.abc import x, y
    >>> u, X, Y = map(Function, 'uXY')

    >>> eq = Eq(D(u(x, y), x, 2), D(u(x, y), y, 2))
    >>> pde_separate_mul(eq, u(x, y), [X(x), Y(y)])
    [Derivative(X(x), (x, 2))/X(x), Derivative(Y(y), (y, 2))/Y(y)]

    """
    return pde_separate(eq, fun, sep, strategy='mul')


def _separate(eq, dep, others):
    """Separate expression into two parts based on dependencies of variables."""

    # FIRST PASS
    # Extract derivatives depending our separable variable...
    terms = set()
    for term in eq.args:
        if term.is_Mul:
            for i in term.args:
                if i.is_Derivative and not i.has(*others):
                    terms.add(term)
                    continue
        elif term.is_Derivative and not term.has(*others):
            terms.add(term)
    # Find the factor that we need to divide by
    div = set()
    for term in terms:
        ext, sep = term.expand().as_independent(dep)
        # Failed?
        if sep.has(*others):
            return None
        div.add(ext)
    # FIXME: Find lcm() of all the divisors and divide with it, instead of
    # current hack :(
    # https://github.com/sympy/sympy/issues/4597
    if len(div) > 0:
        # double sum required or some tests will fail
        eq = Add(*[simplify(Add(*[term/i for i in div])) for term in eq.args])
    # SECOND PASS - separate the derivatives
    div = set()
    lhs = rhs = 0
    for term in eq.args:
        # Check, whether we have already term with independent variable...
        if not term.has(*others):
            lhs += term
            continue
        # ...otherwise, try to separate
        temp, sep = term.expand().as_independent(dep)
        # Failed?
        if sep.has(*others):
            return None
        # Extract the divisors
        div.add(sep)
        rhs -= term.expand()
    # Do the division
    fulldiv = reduce(operator.add, div)
    lhs = simplify(lhs/fulldiv).expand()
    rhs = simplify(rhs/fulldiv).expand()
    # ...and check whether we were successful :)
    if lhs.has(*others) or rhs.has(dep):
        return None
    return [lhs, rhs]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\pivot\cache.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Typed,
    Bool,
    Float,
    Set,
    NoneSet,
    String,
    Integer,
    DateTime,
    Sequence,
)

from openpyxl.descriptors.excel import (
    HexBinary,
    ExtensionList,
    Relation,
)
from openpyxl.descriptors.nested import NestedInteger
from openpyxl.descriptors.sequence import (
    NestedSequence,
    MultiSequence,
    MultiSequencePart,
)
from openpyxl.xml.constants import SHEET_MAIN_NS
from openpyxl.xml.functions import tostring
from openpyxl.packaging.relationship import (
    RelationshipList,
    Relationship,
    get_rels_path
)

from .table import (
    PivotArea,
    Reference,
)
from .fields import (
    Boolean,
    Error,
    Missing,
    Number,
    Text,
    TupleList,
    DateTimeField,
)

class MeasureDimensionMap(Serialisable):

    tagname = "map"

    measureGroup = Integer(allow_none=True)
    dimension = Integer(allow_none=True)

    def __init__(self,
                 measureGroup=None,
                 dimension=None,
                ):
        self.measureGroup = measureGroup
        self.dimension = dimension


class MeasureGroup(Serialisable):

    tagname = "measureGroup"

    name = String()
    caption = String()

    def __init__(self,
                 name=None,
                 caption=None,
                ):
        self.name = name
        self.caption = caption


class PivotDimension(Serialisable):

    tagname = "dimension"

    measure = Bool()
    name = String()
    uniqueName = String()
    caption = String()

    def __init__(self,
                 measure=None,
                 name=None,
                 uniqueName=None,
                 caption=None,
                ):
        self.measure = measure
        self.name = name
        self.uniqueName = uniqueName
        self.caption = caption


class CalculatedMember(Serialisable):

    tagname = "calculatedMember"

    name = String()
    mdx = String()
    memberName = String(allow_none=True)
    hierarchy = String(allow_none=True)
    parent = String(allow_none=True)
    solveOrder = Integer(allow_none=True)
    set = Bool()
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ()

    def __init__(self,
                 name=None,
                 mdx=None,
                 memberName=None,
                 hierarchy=None,
                 parent=None,
                 solveOrder=None,
                 set=None,
                 extLst=None,
                ):
        self.name = name
        self.mdx = mdx
        self.memberName = memberName
        self.hierarchy = hierarchy
        self.parent = parent
        self.solveOrder = solveOrder
        self.set = set
        #self.extLst = extLst


class CalculatedItem(Serialisable):

    tagname = "calculatedItem"

    field = Integer(allow_none=True)
    formula = String()
    pivotArea = Typed(expected_type=PivotArea, )
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('pivotArea', 'extLst')

    def __init__(self,
                 field=None,
                 formula=None,
                 pivotArea=None,
                 extLst=None,
                ):
        self.field = field
        self.formula = formula
        self.pivotArea = pivotArea
        self.extLst = extLst


class ServerFormat(Serialisable):

    tagname = "serverFormat"

    culture = String(allow_none=True)
    format = String(allow_none=True)

    def __init__(self,
                 culture=None,
                 format=None,
                ):
        self.culture = culture
        self.format = format


class Query(Serialisable):

    tagname = "query"

    mdx = String()
    tpls = Typed(expected_type=TupleList, allow_none=True)

    __elements__ = ('tpls',)

    def __init__(self,
                 mdx=None,
                 tpls=None,
                ):
        self.mdx = mdx
        self.tpls = tpls


class OLAPSet(Serialisable):

    tagname = "set"

    count = Integer()
    maxRank = Integer()
    setDefinition = String()
    sortType = NoneSet(values=(['ascending', 'descending', 'ascendingAlpha',
                                'descendingAlpha', 'ascendingNatural', 'descendingNatural']))
    queryFailed = Bool()
    tpls = Typed(expected_type=TupleList, allow_none=True)
    sortByTuple = Typed(expected_type=TupleList, allow_none=True)

    __elements__ = ('tpls', 'sortByTuple')

    def __init__(self,
                 count=None,
                 maxRank=None,
                 setDefinition=None,
                 sortType=None,
                 queryFailed=None,
                 tpls=None,
                 sortByTuple=None,
                ):
        self.count = count
        self.maxRank = maxRank
        self.setDefinition = setDefinition
        self.sortType = sortType
        self.queryFailed = queryFailed
        self.tpls = tpls
        self.sortByTuple = sortByTuple


class PCDSDTCEntries(Serialisable):
    # Implements CT_PCDSDTCEntries

    tagname = "entries"

    count = Integer(allow_none=True)
    # elements are choice
    m = Typed(expected_type=Missing, allow_none=True)
    n = Typed(expected_type=Number, allow_none=True)
    e = Typed(expected_type=Error, allow_none=True)
    s = Typed(expected_type=Text, allow_none=True)

    __elements__ = ('m', 'n', 'e', 's')

    def __init__(self,
                 count=None,
                 m=None,
                 n=None,
                 e=None,
                 s=None,
                ):
        self.count = count
        self.m = m
        self.n = n
        self.e = e
        self.s = s


class TupleCache(Serialisable):

    tagname = "tupleCache"

    entries = Typed(expected_type=PCDSDTCEntries, allow_none=True)
    sets = NestedSequence(expected_type=OLAPSet, count=True)
    queryCache = NestedSequence(expected_type=Query, count=True)
    serverFormats = NestedSequence(expected_type=ServerFormat, count=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('entries', 'sets', 'queryCache', 'serverFormats', 'extLst')

    def __init__(self,
                 entries=None,
                 sets=(),
                 queryCache=(),
                 serverFormats=(),
                 extLst=None,
                ):
        self.entries = entries
        self.sets = sets
        self.queryCache = queryCache
        self.serverFormats = serverFormats
        self.extLst = extLst


class OLAPKPI(Serialisable):

    tagname = "kpi"

    uniqueName = String()
    caption = String(allow_none=True)
    displayFolder = String(allow_none=True)
    measureGroup = String(allow_none=True)
    parent = String(allow_none=True)
    value = String()
    goal = String(allow_none=True)
    status = String(allow_none=True)
    trend = String(allow_none=True)
    weight = String(allow_none=True)
    time = String(allow_none=True)

    def __init__(self,
                 uniqueName=None,
                 caption=None,
                 displayFolder=None,
                 measureGroup=None,
                 parent=None,
                 value=None,
                 goal=None,
                 status=None,
                 trend=None,
                 weight=None,
                 time=None,
                ):
        self.uniqueName = uniqueName
        self.caption = caption
        self.displayFolder = displayFolder
        self.measureGroup = measureGroup
        self.parent = parent
        self.value = value
        self.goal = goal
        self.status = status
        self.trend = trend
        self.weight = weight
        self.time = time


class GroupMember(Serialisable):

    tagname = "groupMember"

    uniqueName = String()
    group = Bool()

    def __init__(self,
                 uniqueName=None,
                 group=None,
                ):
        self.uniqueName = uniqueName
        self.group = group


class LevelGroup(Serialisable):

    tagname = "group"

    name = String()
    uniqueName = String()
    caption = String()
    uniqueParent = String()
    id = Integer()
    groupMembers = NestedSequence(expected_type=GroupMember, count=True)

    __elements__ = ('groupMembers',)

    def __init__(self,
                 name=None,
                 uniqueName=None,
                 caption=None,
                 uniqueParent=None,
                 id=None,
                 groupMembers=(),
                ):
        self.name = name
        self.uniqueName = uniqueName
        self.caption = caption
        self.uniqueParent = uniqueParent
        self.id = id
        self.groupMembers = groupMembers


class GroupLevel(Serialisable):

    tagname = "groupLevel"

    uniqueName = String()
    caption = String()
    user = Bool()
    customRollUp = Bool()
    groups = NestedSequence(expected_type=LevelGroup, count=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('groups', 'extLst')

    def __init__(self,
                 uniqueName=None,
                 caption=None,
                 user=None,
                 customRollUp=None,
                 groups=(),
                 extLst=None,
                ):
        self.uniqueName = uniqueName
        self.caption = caption
        self.user = user
        self.customRollUp = customRollUp
        self.groups = groups
        self.extLst = extLst


class FieldUsage(Serialisable):

    tagname = "fieldUsage"

    x = Integer()

    def __init__(self,
                 x=None,
                ):
        self.x = x


class CacheHierarchy(Serialisable):

    tagname = "cacheHierarchy"

    uniqueName = String()
    caption = String(allow_none=True)
    measure = Bool()
    set = Bool()
    parentSet = Integer(allow_none=True)
    iconSet = Integer()
    attribute = Bool()
    time = Bool()
    keyAttribute = Bool()
    defaultMemberUniqueName = String(allow_none=True)
    allUniqueName = String(allow_none=True)
    allCaption = String(allow_none=True)
    dimensionUniqueName = String(allow_none=True)
    displayFolder = String(allow_none=True)
    measureGroup = String(allow_none=True)
    measures = Bool()
    count = Integer()
    oneField = Bool()
    memberValueDatatype = Integer(allow_none=True)
    unbalanced = Bool(allow_none=True)
    unbalancedGroup = Bool(allow_none=True)
    hidden = Bool()
    fieldsUsage = NestedSequence(expected_type=FieldUsage, count=True)
    groupLevels = NestedSequence(expected_type=GroupLevel, count=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('fieldsUsage', 'groupLevels')

    def __init__(self,
                 uniqueName="",
                 caption=None,
                 measure=None,
                 set=None,
                 parentSet=None,
                 iconSet=0,
                 attribute=None,
                 time=None,
                 keyAttribute=None,
                 defaultMemberUniqueName=None,
                 allUniqueName=None,
                 allCaption=None,
                 dimensionUniqueName=None,
                 displayFolder=None,
                 measureGroup=None,
                 measures=None,
                 count=None,
                 oneField=None,
                 memberValueDatatype=None,
                 unbalanced=None,
                 unbalancedGroup=None,
                 hidden=None,
                 fieldsUsage=(),
                 groupLevels=(),
                 extLst=None,
                ):
        self.uniqueName = uniqueName
        self.caption = caption
        self.measure = measure
        self.set = set
        self.parentSet = parentSet
        self.iconSet = iconSet
        self.attribute = attribute
        self.time = time
        self.keyAttribute = keyAttribute
        self.defaultMemberUniqueName = defaultMemberUniqueName
        self.allUniqueName = allUniqueName
        self.allCaption = allCaption
        self.dimensionUniqueName = dimensionUniqueName
        self.displayFolder = displayFolder
        self.measureGroup = measureGroup
        self.measures = measures
        self.count = count
        self.oneField = oneField
        self.memberValueDatatype = memberValueDatatype
        self.unbalanced = unbalanced
        self.unbalancedGroup = unbalancedGroup
        self.hidden = hidden
        self.fieldsUsage = fieldsUsage
        self.groupLevels = groupLevels
        self.extLst = extLst


class GroupItems(Serialisable):

    tagname = "groupItems"

    m = Sequence(expected_type=Missing)
    n = Sequence(expected_type=Number)
    b = Sequence(expected_type=Boolean)
    e = Sequence(expected_type=Error)
    s = Sequence(expected_type=Text)
    d = Sequence(expected_type=DateTimeField,)

    __elements__ = ('m', 'n', 'b', 'e', 's', 'd')
    __attrs__ = ("count", )

    def __init__(self,
                 count=None,
                 m=(),
                 n=(),
                 b=(),
                 e=(),
                 s=(),
                 d=(),
                ):
        self.m = m
        self.n = n
        self.b = b
        self.e = e
        self.s = s
        self.d = d


    @property
    def count(self):
        return len(self.m + self.n + self.b + self.e + self.s + self.d)


class RangePr(Serialisable):

    tagname = "rangePr"

    autoStart = Bool(allow_none=True)
    autoEnd = Bool(allow_none=True)
    groupBy = NoneSet(values=(['range', 'seconds', 'minutes', 'hours', 'days',
                           'months', 'quarters', 'years']))
    startNum = Float(allow_none=True)
    endNum = Float(allow_none=True)
    startDate = DateTime(allow_none=True)
    endDate = DateTime(allow_none=True)
    groupInterval = Float(allow_none=True)

    def __init__(self,
                 autoStart=True,
                 autoEnd=True,
                 groupBy="range",
                 startNum=None,
                 endNum=None,
                 startDate=None,
                 endDate=None,
                 groupInterval=1,
                ):
        self.autoStart = autoStart
        self.autoEnd = autoEnd
        self.groupBy = groupBy
        self.startNum = startNum
        self.endNum = endNum
        self.startDate = startDate
        self.endDate = endDate
        self.groupInterval = groupInterval


class FieldGroup(Serialisable):

    tagname = "fieldGroup"

    par = Integer(allow_none=True)
    base = Integer(allow_none=True)
    rangePr = Typed(expected_type=RangePr, allow_none=True)
    discretePr = NestedSequence(expected_type=NestedInteger, count=True)
    groupItems = Typed(expected_type=GroupItems, allow_none=True)

    __elements__ = ('rangePr', 'discretePr', 'groupItems')

    def __init__(self,
                 par=None,
                 base=None,
                 rangePr=None,
                 discretePr=(),
                 groupItems=None,
                ):
        self.par = par
        self.base = base
        self.rangePr = rangePr
        self.discretePr = discretePr
        self.groupItems = groupItems


class SharedItems(Serialisable):

    tagname = "sharedItems"

    _fields = MultiSequence()
    m = MultiSequencePart(expected_type=Missing, store="_fields")
    n = MultiSequencePart(expected_type=Number, store="_fields")
    b = MultiSequencePart(expected_type=Boolean, store="_fields")
    e = MultiSequencePart(expected_type=Error, store="_fields")
    s = MultiSequencePart(expected_type=Text,  store="_fields")
    d = MultiSequencePart(expected_type=DateTimeField, store="_fields")
    # attributes are optional and must be derived from associated cache records
    containsSemiMixedTypes = Bool(allow_none=True)
    containsNonDate = Bool(allow_none=True)
    containsDate = Bool(allow_none=True)
    containsString = Bool(allow_none=True)
    containsBlank = Bool(allow_none=True)
    containsMixedTypes = Bool(allow_none=True)
    containsNumber = Bool(allow_none=True)
    containsInteger = Bool(allow_none=True)
    minValue = Float(allow_none=True)
    maxValue = Float(allow_none=True)
    minDate = DateTime(allow_none=True)
    maxDate = DateTime(allow_none=True)
    longText = Bool(allow_none=True)

    __attrs__ = ('count', 'containsBlank', 'containsDate', 'containsInteger',
                 'containsMixedTypes', 'containsNonDate', 'containsNumber',
                 'containsSemiMixedTypes', 'containsString', 'minValue', 'maxValue',
                 'minDate', 'maxDate', 'longText')

    def __init__(self,
                 _fields=(),
                 containsSemiMixedTypes=None,
                 containsNonDate=None,
                 containsDate=None,
                 containsString=None,
                 containsBlank=None,
                 containsMixedTypes=None,
                 containsNumber=None,
                 containsInteger=None,
                 minValue=None,
                 maxValue=None,
                 minDate=None,
                 maxDate=None,
                 count=None,
                 longText=None,
                ):
        self._fields = _fields
        self.containsBlank = containsBlank
        self.containsDate = containsDate
        self.containsNonDate = containsNonDate
        self.containsString = containsString
        self.containsMixedTypes = containsMixedTypes
        self.containsSemiMixedTypes = containsSemiMixedTypes
        self.containsNumber = containsNumber
        self.containsInteger = containsInteger
        self.minValue = minValue
        self.maxValue = maxValue
        self.minDate = minDate
        self.maxDate = maxDate
        self.longText = longText


    @property
    def count(self):
        return len(self._fields)


class CacheField(Serialisable):

    tagname = "cacheField"

    sharedItems = Typed(expected_type=SharedItems, allow_none=True)
    fieldGroup = Typed(expected_type=FieldGroup, allow_none=True)
    mpMap = NestedInteger(allow_none=True, attribute="v")
    extLst = Typed(expected_type=ExtensionList, allow_none=True)
    name = String()
    caption = String(allow_none=True)
    propertyName = String(allow_none=True)
    serverField = Bool(allow_none=True)
    uniqueList = Bool(allow_none=True)
    numFmtId = Integer(allow_none=True)
    formula = String(allow_none=True)
    sqlType = Integer(allow_none=True)
    hierarchy = Integer(allow_none=True)
    level = Integer(allow_none=True)
    databaseField = Bool(allow_none=True)
    mappingCount = Integer(allow_none=True)
    memberPropertyField = Bool(allow_none=True)

    __elements__ = ('sharedItems', 'fieldGroup', 'mpMap')

    def __init__(self,
                 sharedItems=None,
                 fieldGroup=None,
                 mpMap=None,
                 extLst=None,
                 name=None,
                 caption=None,
                 propertyName=None,
                 serverField=None,
                 uniqueList=True,
                 numFmtId=None,
                 formula=None,
                 sqlType=0,
                 hierarchy=0,
                 level=0,
                 databaseField=True,
                 mappingCount=None,
                 memberPropertyField=None,
                ):
        self.sharedItems = sharedItems
        self.fieldGroup = fieldGroup
        self.mpMap = mpMap
        self.extLst = extLst
        self.name = name
        self.caption = caption
        self.propertyName = propertyName
        self.serverField = serverField
        self.uniqueList = uniqueList
        self.numFmtId = numFmtId
        self.formula = formula
        self.sqlType = sqlType
        self.hierarchy = hierarchy
        self.level = level
        self.databaseField = databaseField
        self.mappingCount = mappingCount
        self.memberPropertyField = memberPropertyField


class RangeSet(Serialisable):

    tagname = "rangeSet"

    i1 = Integer(allow_none=True)
    i2 = Integer(allow_none=True)
    i3 = Integer(allow_none=True)
    i4 = Integer(allow_none=True)
    ref = String()
    name = String(allow_none=True)
    sheet = String(allow_none=True)

    def __init__(self,
                 i1=None,
                 i2=None,
                 i3=None,
                 i4=None,
                 ref=None,
                 name=None,
                 sheet=None,
                ):
        self.i1 = i1
        self.i2 = i2
        self.i3 = i3
        self.i4 = i4
        self.ref = ref
        self.name = name
        self.sheet = sheet


class PageItem(Serialisable):

    tagname = "pageItem"

    name = String()

    def __init__(self,
                 name=None,
                ):
        self.name = name


class Consolidation(Serialisable):

    tagname = "consolidation"

    autoPage = Bool(allow_none=True)
    pages = NestedSequence(expected_type=PageItem, count=True)
    rangeSets = NestedSequence(expected_type=RangeSet, count=True)

    __elements__ = ('pages', 'rangeSets')

    def __init__(self,
                 autoPage=None,
                 pages=(),
                 rangeSets=(),
                ):
        self.autoPage = autoPage
        self.pages = pages
        self.rangeSets = rangeSets


class WorksheetSource(Serialisable):

    tagname = "worksheetSource"

    ref = String(allow_none=True)
    name = String(allow_none=True)
    sheet = String(allow_none=True)

    def __init__(self,
                 ref=None,
                 name=None,
                 sheet=None,
                ):
        self.ref = ref
        self.name = name
        self.sheet = sheet


class CacheSource(Serialisable):

    tagname = "cacheSource"

    type = Set(values=(['worksheet', 'external', 'consolidation', 'scenario']))
    connectionId = Integer(allow_none=True)
    # some elements are choice
    worksheetSource = Typed(expected_type=WorksheetSource, allow_none=True)
    consolidation = Typed(expected_type=Consolidation, allow_none=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('worksheetSource', 'consolidation',)

    def __init__(self,
                 type=None,
                 connectionId=None,
                 worksheetSource=None,
                 consolidation=None,
                 extLst=None,
                ):
        self.type = type
        self.connectionId = connectionId
        self.worksheetSource = worksheetSource
        self.consolidation = consolidation


class CacheDefinition(Serialisable):

    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"
    rel_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition"
    _id = 1
    _path = "/xl/pivotCache/pivotCacheDefinition{0}.xml"
    records = None

    tagname = "pivotCacheDefinition"

    invalid = Bool(allow_none=True)
    saveData = Bool(allow_none=True)
    refreshOnLoad = Bool(allow_none=True)
    optimizeMemory = Bool(allow_none=True)
    enableRefresh = Bool(allow_none=True)
    refreshedBy = String(allow_none=True)
    refreshedDate = Float(allow_none=True)
    refreshedDateIso = DateTime(allow_none=True)
    backgroundQuery = Bool(allow_none=True)
    missingItemsLimit = Integer(allow_none=True)
    createdVersion = Integer(allow_none=True)
    refreshedVersion = Integer(allow_none=True)
    minRefreshableVersion = Integer(allow_none=True)
    recordCount = Integer(allow_none=True)
    upgradeOnRefresh = Bool(allow_none=True)
    supportSubquery = Bool(allow_none=True)
    supportAdvancedDrill = Bool(allow_none=True)
    cacheSource = Typed(expected_type=CacheSource)
    cacheFields = NestedSequence(expected_type=CacheField, count=True)
    cacheHierarchies = NestedSequence(expected_type=CacheHierarchy, allow_none=True)
    kpis = NestedSequence(expected_type=OLAPKPI, count=True)
    tupleCache = Typed(expected_type=TupleCache, allow_none=True)
    calculatedItems = NestedSequence(expected_type=CalculatedItem, count=True)
    calculatedMembers = NestedSequence(expected_type=CalculatedMember, count=True)
    dimensions = NestedSequence(expected_type=PivotDimension, allow_none=True)
    measureGroups = NestedSequence(expected_type=MeasureGroup, count=True)
    maps = NestedSequence(expected_type=MeasureDimensionMap, count=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)
    id = Relation()

    __elements__ = ('cacheSource', 'cacheFields', 'cacheHierarchies', 'kpis',
                    'tupleCache', 'calculatedItems', 'calculatedMembers', 'dimensions',
                    'measureGroups', 'maps',)

    def __init__(self,
                 invalid=None,
                 saveData=None,
                 refreshOnLoad=None,
                 optimizeMemory=None,
                 enableRefresh=None,
                 refreshedBy=None,
                 refreshedDate=None,
                 refreshedDateIso=None,
                 backgroundQuery=None,
                 missingItemsLimit=None,
                 createdVersion=None,
                 refreshedVersion=None,
                 minRefreshableVersion=None,
                 recordCount=None,
                 upgradeOnRefresh=None,
                 tupleCache=None,
                 supportSubquery=None,
                 supportAdvancedDrill=None,
                 cacheSource=None,
                 cacheFields=(),
                 cacheHierarchies=(),
                 kpis=(),
                 calculatedItems=(),
                 calculatedMembers=(),
                 dimensions=(),
                 measureGroups=(),
                 maps=(),
                 extLst=None,
                 id = None,
                ):
        self.invalid = invalid
        self.saveData = saveData
        self.refreshOnLoad = refreshOnLoad
        self.optimizeMemory = optimizeMemory
        self.enableRefresh = enableRefresh
        self.refreshedBy = refreshedBy
        self.refreshedDate = refreshedDate
        self.refreshedDateIso = refreshedDateIso
        self.backgroundQuery = backgroundQuery
        self.missingItemsLimit = missingItemsLimit
        self.createdVersion = createdVersion
        self.refreshedVersion = refreshedVersion
        self.minRefreshableVersion = minRefreshableVersion
        self.recordCount = recordCount
        self.upgradeOnRefresh = upgradeOnRefresh
        self.supportSubquery = supportSubquery
        self.supportAdvancedDrill = supportAdvancedDrill
        self.cacheSource = cacheSource
        self.cacheFields = cacheFields
        self.cacheHierarchies = cacheHierarchies
        self.kpis = kpis
        self.tupleCache = tupleCache
        self.calculatedItems = calculatedItems
        self.calculatedMembers = calculatedMembers
        self.dimensions = dimensions
        self.measureGroups = measureGroups
        self.maps = maps
        self.id = id


    def to_tree(self):
        node = super().to_tree()
        node.set("xmlns", SHEET_MAIN_NS)
        return node


    @property
    def path(self):
        return self._path.format(self._id)


    def _write(self, archive, manifest):
        """
        Add to zipfile and update manifest
        """
        self._write_rels(archive, manifest)
        xml = tostring(self.to_tree())
        archive.writestr(self.path[1:], xml)
        manifest.append(self)


    def _write_rels(self, archive, manifest):
        """
        Write the relevant child objects and add links
        """
        if self.records is None:
            return

        rels = RelationshipList()
        r = Relationship(Type=self.records.rel_type, Target=self.records.path)
        rels.append(r)
        self.id = r.id
        self.records._id = self._id
        self.records._write(archive, manifest)

        path = get_rels_path(self.path)
        xml = tostring(rels.to_tree())
        archive.writestr(path[1:], xml)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\engine\events.py ===
# engine/events.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php


from __future__ import annotations

import typing
from typing import Any
from typing import Dict
from typing import Optional
from typing import Tuple
from typing import Type
from typing import Union

from .base import Connection
from .base import Engine
from .interfaces import ConnectionEventsTarget
from .interfaces import DBAPIConnection
from .interfaces import DBAPICursor
from .interfaces import Dialect
from .. import event
from .. import exc
from ..util.typing import Literal

if typing.TYPE_CHECKING:
    from .interfaces import _CoreMultiExecuteParams
    from .interfaces import _CoreSingleExecuteParams
    from .interfaces import _DBAPIAnyExecuteParams
    from .interfaces import _DBAPIMultiExecuteParams
    from .interfaces import _DBAPISingleExecuteParams
    from .interfaces import _ExecuteOptions
    from .interfaces import ExceptionContext
    from .interfaces import ExecutionContext
    from .result import Result
    from ..pool import ConnectionPoolEntry
    from ..sql import Executable
    from ..sql.elements import BindParameter


class ConnectionEvents(event.Events[ConnectionEventsTarget]):
    """Available events for
    :class:`_engine.Connection` and :class:`_engine.Engine`.

    The methods here define the name of an event as well as the names of
    members that are passed to listener functions.

    An event listener can be associated with any
    :class:`_engine.Connection` or :class:`_engine.Engine`
    class or instance, such as an :class:`_engine.Engine`, e.g.::

        from sqlalchemy import event, create_engine


        def before_cursor_execute(
            conn, cursor, statement, parameters, context, executemany
        ):
            log.info("Received statement: %s", statement)


        engine = create_engine("postgresql+psycopg2://scott:tiger@localhost/test")
        event.listen(engine, "before_cursor_execute", before_cursor_execute)

    or with a specific :class:`_engine.Connection`::

        with engine.begin() as conn:

            @event.listens_for(conn, "before_cursor_execute")
            def before_cursor_execute(
                conn, cursor, statement, parameters, context, executemany
            ):
                log.info("Received statement: %s", statement)

    When the methods are called with a `statement` parameter, such as in
    :meth:`.after_cursor_execute` or :meth:`.before_cursor_execute`,
    the statement is the exact SQL string that was prepared for transmission
    to the DBAPI ``cursor`` in the connection's :class:`.Dialect`.

    The :meth:`.before_execute` and :meth:`.before_cursor_execute`
    events can also be established with the ``retval=True`` flag, which
    allows modification of the statement and parameters to be sent
    to the database.  The :meth:`.before_cursor_execute` event is
    particularly useful here to add ad-hoc string transformations, such
    as comments, to all executions::

        from sqlalchemy.engine import Engine
        from sqlalchemy import event


        @event.listens_for(Engine, "before_cursor_execute", retval=True)
        def comment_sql_calls(
            conn, cursor, statement, parameters, context, executemany
        ):
            statement = statement + " -- some comment"
            return statement, parameters

    .. note:: :class:`_events.ConnectionEvents` can be established on any
       combination of :class:`_engine.Engine`, :class:`_engine.Connection`,
       as well
       as instances of each of those classes.  Events across all
       four scopes will fire off for a given instance of
       :class:`_engine.Connection`.  However, for performance reasons, the
       :class:`_engine.Connection` object determines at instantiation time
       whether or not its parent :class:`_engine.Engine` has event listeners
       established.   Event listeners added to the :class:`_engine.Engine`
       class or to an instance of :class:`_engine.Engine`
       *after* the instantiation
       of a dependent :class:`_engine.Connection` instance will usually
       *not* be available on that :class:`_engine.Connection` instance.
       The newly
       added listeners will instead take effect for
       :class:`_engine.Connection`
       instances created subsequent to those event listeners being
       established on the parent :class:`_engine.Engine` class or instance.

    :param retval=False: Applies to the :meth:`.before_execute` and
      :meth:`.before_cursor_execute` events only.  When True, the
      user-defined event function must have a return value, which
      is a tuple of parameters that replace the given statement
      and parameters.  See those methods for a description of
      specific return arguments.

    """  # noqa

    _target_class_doc = "SomeEngine"
    _dispatch_target = ConnectionEventsTarget

    @classmethod
    def _accept_with(
        cls,
        target: Union[ConnectionEventsTarget, Type[ConnectionEventsTarget]],
        identifier: str,
    ) -> Optional[Union[ConnectionEventsTarget, Type[ConnectionEventsTarget]]]:
        default_dispatch = super()._accept_with(target, identifier)
        if default_dispatch is None and hasattr(
            target, "_no_async_engine_events"
        ):
            target._no_async_engine_events()

        return default_dispatch

    @classmethod
    def _listen(
        cls,
        event_key: event._EventKey[ConnectionEventsTarget],
        *,
        retval: bool = False,
        **kw: Any,
    ) -> None:
        target, identifier, fn = (
            event_key.dispatch_target,
            event_key.identifier,
            event_key._listen_fn,
        )
        target._has_events = True

        if not retval:
            if identifier == "before_execute":
                orig_fn = fn

                def wrap_before_execute(  # type: ignore
                    conn, clauseelement, multiparams, params, execution_options
                ):
                    orig_fn(
                        conn,
                        clauseelement,
                        multiparams,
                        params,
                        execution_options,
                    )
                    return clauseelement, multiparams, params

                fn = wrap_before_execute
            elif identifier == "before_cursor_execute":
                orig_fn = fn

                def wrap_before_cursor_execute(  # type: ignore
                    conn, cursor, statement, parameters, context, executemany
                ):
                    orig_fn(
                        conn,
                        cursor,
                        statement,
                        parameters,
                        context,
                        executemany,
                    )
                    return statement, parameters

                fn = wrap_before_cursor_execute
        elif retval and identifier not in (
            "before_execute",
            "before_cursor_execute",
        ):
            raise exc.ArgumentError(
                "Only the 'before_execute', "
                "'before_cursor_execute' and 'handle_error' engine "
                "event listeners accept the 'retval=True' "
                "argument."
            )
        event_key.with_wrapper(fn).base_listen()

    @event._legacy_signature(
        "1.4",
        ["conn", "clauseelement", "multiparams", "params"],
        lambda conn, clauseelement, multiparams, params, execution_options: (
            conn,
            clauseelement,
            multiparams,
            params,
        ),
    )
    def before_execute(
        self,
        conn: Connection,
        clauseelement: Executable,
        multiparams: _CoreMultiExecuteParams,
        params: _CoreSingleExecuteParams,
        execution_options: _ExecuteOptions,
    ) -> Optional[
        Tuple[Executable, _CoreMultiExecuteParams, _CoreSingleExecuteParams]
    ]:
        """Intercept high level execute() events, receiving uncompiled
        SQL constructs and other objects prior to rendering into SQL.

        This event is good for debugging SQL compilation issues as well
        as early manipulation of the parameters being sent to the database,
        as the parameter lists will be in a consistent format here.

        This event can be optionally established with the ``retval=True``
        flag.  The ``clauseelement``, ``multiparams``, and ``params``
        arguments should be returned as a three-tuple in this case::

            @event.listens_for(Engine, "before_execute", retval=True)
            def before_execute(conn, clauseelement, multiparams, params):
                # do something with clauseelement, multiparams, params
                return clauseelement, multiparams, params

        :param conn: :class:`_engine.Connection` object
        :param clauseelement: SQL expression construct, :class:`.Compiled`
         instance, or string statement passed to
         :meth:`_engine.Connection.execute`.
        :param multiparams: Multiple parameter sets, a list of dictionaries.
        :param params: Single parameter set, a single dictionary.
        :param execution_options: dictionary of execution
         options passed along with the statement, if any.  This is a merge
         of all options that will be used, including those of the statement,
         the connection, and those passed in to the method itself for
         the 2.0 style of execution.

         .. versionadded: 1.4

        .. seealso::

            :meth:`.before_cursor_execute`

        """

    @event._legacy_signature(
        "1.4",
        ["conn", "clauseelement", "multiparams", "params", "result"],
        lambda conn, clauseelement, multiparams, params, execution_options, result: (  # noqa
            conn,
            clauseelement,
            multiparams,
            params,
            result,
        ),
    )
    def after_execute(
        self,
        conn: Connection,
        clauseelement: Executable,
        multiparams: _CoreMultiExecuteParams,
        params: _CoreSingleExecuteParams,
        execution_options: _ExecuteOptions,
        result: Result[Any],
    ) -> None:
        """Intercept high level execute() events after execute.


        :param conn: :class:`_engine.Connection` object
        :param clauseelement: SQL expression construct, :class:`.Compiled`
         instance, or string statement passed to
         :meth:`_engine.Connection.execute`.
        :param multiparams: Multiple parameter sets, a list of dictionaries.
        :param params: Single parameter set, a single dictionary.
        :param execution_options: dictionary of execution
         options passed along with the statement, if any.  This is a merge
         of all options that will be used, including those of the statement,
         the connection, and those passed in to the method itself for
         the 2.0 style of execution.

         .. versionadded: 1.4

        :param result: :class:`_engine.CursorResult` generated by the
         execution.

        """

    def before_cursor_execute(
        self,
        conn: Connection,
        cursor: DBAPICursor,
        statement: str,
        parameters: _DBAPIAnyExecuteParams,
        context: Optional[ExecutionContext],
        executemany: bool,
    ) -> Optional[Tuple[str, _DBAPIAnyExecuteParams]]:
        """Intercept low-level cursor execute() events before execution,
        receiving the string SQL statement and DBAPI-specific parameter list to
        be invoked against a cursor.

        This event is a good choice for logging as well as late modifications
        to the SQL string.  It's less ideal for parameter modifications except
        for those which are specific to a target backend.

        This event can be optionally established with the ``retval=True``
        flag.  The ``statement`` and ``parameters`` arguments should be
        returned as a two-tuple in this case::

            @event.listens_for(Engine, "before_cursor_execute", retval=True)
            def before_cursor_execute(
                conn, cursor, statement, parameters, context, executemany
            ):
                # do something with statement, parameters
                return statement, parameters

        See the example at :class:`_events.ConnectionEvents`.

        :param conn: :class:`_engine.Connection` object
        :param cursor: DBAPI cursor object
        :param statement: string SQL statement, as to be passed to the DBAPI
        :param parameters: Dictionary, tuple, or list of parameters being
         passed to the ``execute()`` or ``executemany()`` method of the
         DBAPI ``cursor``.  In some cases may be ``None``.
        :param context: :class:`.ExecutionContext` object in use.  May
         be ``None``.
        :param executemany: boolean, if ``True``, this is an ``executemany()``
         call, if ``False``, this is an ``execute()`` call.

        .. seealso::

            :meth:`.before_execute`

            :meth:`.after_cursor_execute`

        """

    def after_cursor_execute(
        self,
        conn: Connection,
        cursor: DBAPICursor,
        statement: str,
        parameters: _DBAPIAnyExecuteParams,
        context: Optional[ExecutionContext],
        executemany: bool,
    ) -> None:
        """Intercept low-level cursor execute() events after execution.

        :param conn: :class:`_engine.Connection` object
        :param cursor: DBAPI cursor object.  Will have results pending
         if the statement was a SELECT, but these should not be consumed
         as they will be needed by the :class:`_engine.CursorResult`.
        :param statement: string SQL statement, as passed to the DBAPI
        :param parameters: Dictionary, tuple, or list of parameters being
         passed to the ``execute()`` or ``executemany()`` method of the
         DBAPI ``cursor``.  In some cases may be ``None``.
        :param context: :class:`.ExecutionContext` object in use.  May
         be ``None``.
        :param executemany: boolean, if ``True``, this is an ``executemany()``
         call, if ``False``, this is an ``execute()`` call.

        """

    @event._legacy_signature(
        "2.0", ["conn", "branch"], converter=lambda conn: (conn, False)
    )
    def engine_connect(self, conn: Connection) -> None:
        """Intercept the creation of a new :class:`_engine.Connection`.

        This event is called typically as the direct result of calling
        the :meth:`_engine.Engine.connect` method.

        It differs from the :meth:`_events.PoolEvents.connect` method, which
        refers to the actual connection to a database at the DBAPI level;
        a DBAPI connection may be pooled and reused for many operations.
        In contrast, this event refers only to the production of a higher level
        :class:`_engine.Connection` wrapper around such a DBAPI connection.

        It also differs from the :meth:`_events.PoolEvents.checkout` event
        in that it is specific to the :class:`_engine.Connection` object,
        not the
        DBAPI connection that :meth:`_events.PoolEvents.checkout` deals with,
        although
        this DBAPI connection is available here via the
        :attr:`_engine.Connection.connection` attribute.
        But note there can in fact
        be multiple :meth:`_events.PoolEvents.checkout`
        events within the lifespan
        of a single :class:`_engine.Connection` object, if that
        :class:`_engine.Connection`
        is invalidated and re-established.

        :param conn: :class:`_engine.Connection` object.

        .. seealso::

            :meth:`_events.PoolEvents.checkout`
            the lower-level pool checkout event
            for an individual DBAPI connection

        """

    def set_connection_execution_options(
        self, conn: Connection, opts: Dict[str, Any]
    ) -> None:
        """Intercept when the :meth:`_engine.Connection.execution_options`
        method is called.

        This method is called after the new :class:`_engine.Connection`
        has been
        produced, with the newly updated execution options collection, but
        before the :class:`.Dialect` has acted upon any of those new options.

        Note that this method is not called when a new
        :class:`_engine.Connection`
        is produced which is inheriting execution options from its parent
        :class:`_engine.Engine`; to intercept this condition, use the
        :meth:`_events.ConnectionEvents.engine_connect` event.

        :param conn: The newly copied :class:`_engine.Connection` object

        :param opts: dictionary of options that were passed to the
         :meth:`_engine.Connection.execution_options` method.
         This dictionary may be modified in place to affect the ultimate
         options which take effect.

         .. versionadded:: 2.0 the ``opts`` dictionary may be modified
            in place.


        .. seealso::

            :meth:`_events.ConnectionEvents.set_engine_execution_options`
            - event
            which is called when :meth:`_engine.Engine.execution_options`
            is called.


        """

    def set_engine_execution_options(
        self, engine: Engine, opts: Dict[str, Any]
    ) -> None:
        """Intercept when the :meth:`_engine.Engine.execution_options`
        method is called.

        The :meth:`_engine.Engine.execution_options` method produces a shallow
        copy of the :class:`_engine.Engine` which stores the new options.
        That new
        :class:`_engine.Engine` is passed here.
        A particular application of this
        method is to add a :meth:`_events.ConnectionEvents.engine_connect`
        event
        handler to the given :class:`_engine.Engine`
        which will perform some per-
        :class:`_engine.Connection` task specific to these execution options.

        :param conn: The newly copied :class:`_engine.Engine` object

        :param opts: dictionary of options that were passed to the
         :meth:`_engine.Connection.execution_options` method.
         This dictionary may be modified in place to affect the ultimate
         options which take effect.

         .. versionadded:: 2.0 the ``opts`` dictionary may be modified
            in place.

        .. seealso::

            :meth:`_events.ConnectionEvents.set_connection_execution_options`
            - event
            which is called when :meth:`_engine.Connection.execution_options`
            is
            called.

        """

    def engine_disposed(self, engine: Engine) -> None:
        """Intercept when the :meth:`_engine.Engine.dispose` method is called.

        The :meth:`_engine.Engine.dispose` method instructs the engine to
        "dispose" of it's connection pool (e.g. :class:`_pool.Pool`), and
        replaces it with a new one.  Disposing of the old pool has the
        effect that existing checked-in connections are closed.  The new
        pool does not establish any new connections until it is first used.

        This event can be used to indicate that resources related to the
        :class:`_engine.Engine` should also be cleaned up,
        keeping in mind that the
        :class:`_engine.Engine`
        can still be used for new requests in which case
        it re-acquires connection resources.

        """

    def begin(self, conn: Connection) -> None:
        """Intercept begin() events.

        :param conn: :class:`_engine.Connection` object

        """

    def rollback(self, conn: Connection) -> None:
        """Intercept rollback() events, as initiated by a
        :class:`.Transaction`.

        Note that the :class:`_pool.Pool` also "auto-rolls back"
        a DBAPI connection upon checkin, if the ``reset_on_return``
        flag is set to its default value of ``'rollback'``.
        To intercept this
        rollback, use the :meth:`_events.PoolEvents.reset` hook.

        :param conn: :class:`_engine.Connection` object

        .. seealso::

            :meth:`_events.PoolEvents.reset`

        """

    def commit(self, conn: Connection) -> None:
        """Intercept commit() events, as initiated by a
        :class:`.Transaction`.

        Note that the :class:`_pool.Pool` may also "auto-commit"
        a DBAPI connection upon checkin, if the ``reset_on_return``
        flag is set to the value ``'commit'``.  To intercept this
        commit, use the :meth:`_events.PoolEvents.reset` hook.

        :param conn: :class:`_engine.Connection` object
        """

    def savepoint(self, conn: Connection, name: str) -> None:
        """Intercept savepoint() events.

        :param conn: :class:`_engine.Connection` object
        :param name: specified name used for the savepoint.

        """

    def rollback_savepoint(
        self, conn: Connection, name: str, context: None
    ) -> None:
        """Intercept rollback_savepoint() events.

        :param conn: :class:`_engine.Connection` object
        :param name: specified name used for the savepoint.
        :param context: not used

        """
        # TODO: deprecate "context"

    def release_savepoint(
        self, conn: Connection, name: str, context: None
    ) -> None:
        """Intercept release_savepoint() events.

        :param conn: :class:`_engine.Connection` object
        :param name: specified name used for the savepoint.
        :param context: not used

        """
        # TODO: deprecate "context"

    def begin_twophase(self, conn: Connection, xid: Any) -> None:
        """Intercept begin_twophase() events.

        :param conn: :class:`_engine.Connection` object
        :param xid: two-phase XID identifier

        """

    def prepare_twophase(self, conn: Connection, xid: Any) -> None:
        """Intercept prepare_twophase() events.

        :param conn: :class:`_engine.Connection` object
        :param xid: two-phase XID identifier
        """

    def rollback_twophase(
        self, conn: Connection, xid: Any, is_prepared: bool
    ) -> None:
        """Intercept rollback_twophase() events.

        :param conn: :class:`_engine.Connection` object
        :param xid: two-phase XID identifier
        :param is_prepared: boolean, indicates if
         :meth:`.TwoPhaseTransaction.prepare` was called.

        """

    def commit_twophase(
        self, conn: Connection, xid: Any, is_prepared: bool
    ) -> None:
        """Intercept commit_twophase() events.

        :param conn: :class:`_engine.Connection` object
        :param xid: two-phase XID identifier
        :param is_prepared: boolean, indicates if
         :meth:`.TwoPhaseTransaction.prepare` was called.

        """


class DialectEvents(event.Events[Dialect]):
    """event interface for execution-replacement functions.

    These events allow direct instrumentation and replacement
    of key dialect functions which interact with the DBAPI.

    .. note::

        :class:`.DialectEvents` hooks should be considered **semi-public**
        and experimental.
        These hooks are not for general use and are only for those situations
        where intricate re-statement of DBAPI mechanics must be injected onto
        an existing dialect.  For general-use statement-interception events,
        please use the :class:`_events.ConnectionEvents` interface.

    .. seealso::

        :meth:`_events.ConnectionEvents.before_cursor_execute`

        :meth:`_events.ConnectionEvents.before_execute`

        :meth:`_events.ConnectionEvents.after_cursor_execute`

        :meth:`_events.ConnectionEvents.after_execute`

    """

    _target_class_doc = "SomeEngine"
    _dispatch_target = Dialect

    @classmethod
    def _listen(
        cls,
        event_key: event._EventKey[Dialect],
        *,
        retval: bool = False,
        **kw: Any,
    ) -> None:
        target = event_key.dispatch_target

        target._has_events = True
        event_key.base_listen()

    @classmethod
    def _accept_with(
        cls,
        target: Union[Engine, Type[Engine], Dialect, Type[Dialect]],
        identifier: str,
    ) -> Optional[Union[Dialect, Type[Dialect]]]:
        if isinstance(target, type):
            if issubclass(target, Engine):
                return Dialect
            elif issubclass(target, Dialect):
                return target
        elif isinstance(target, Engine):
            return target.dialect
        elif isinstance(target, Dialect):
            return target
        elif isinstance(target, Connection) and identifier == "handle_error":
            raise exc.InvalidRequestError(
                "The handle_error() event hook as of SQLAlchemy 2.0 is "
                "established on the Dialect, and may only be applied to the "
                "Engine as a whole or to a specific Dialect as a whole, "
                "not on a per-Connection basis."
            )
        elif hasattr(target, "_no_async_engine_events"):
            target._no_async_engine_events()
        else:
            return None

    def handle_error(
        self, exception_context: ExceptionContext
    ) -> Optional[BaseException]:
        r"""Intercept all exceptions processed by the
        :class:`_engine.Dialect`, typically but not limited to those
        emitted within the scope of a :class:`_engine.Connection`.

        .. versionchanged:: 2.0 the :meth:`.DialectEvents.handle_error` event
           is moved to the :class:`.DialectEvents` class, moved from the
           :class:`.ConnectionEvents` class, so that it may also participate in
           the "pre ping" operation configured with the
           :paramref:`_sa.create_engine.pool_pre_ping` parameter. The event
           remains registered by using the :class:`_engine.Engine` as the event
           target, however note that using the :class:`_engine.Connection` as
           an event target for :meth:`.DialectEvents.handle_error` is no longer
           supported.

        This includes all exceptions emitted by the DBAPI as well as
        within SQLAlchemy's statement invocation process, including
        encoding errors and other statement validation errors.  Other areas
        in which the event is invoked include transaction begin and end,
        result row fetching, cursor creation.

        Note that :meth:`.handle_error` may support new kinds of exceptions
        and new calling scenarios at *any time*.  Code which uses this
        event must expect new calling patterns to be present in minor
        releases.

        To support the wide variety of members that correspond to an exception,
        as well as to allow extensibility of the event without backwards
        incompatibility, the sole argument received is an instance of
        :class:`.ExceptionContext`.   This object contains data members
        representing detail about the exception.

        Use cases supported by this hook include:

        * read-only, low-level exception handling for logging and
          debugging purposes
        * Establishing whether a DBAPI connection error message indicates
          that the database connection needs to be reconnected, including
          for the "pre_ping" handler used by **some** dialects
        * Establishing or disabling whether a connection or the owning
          connection pool is invalidated or expired in response to a
          specific exception
        * exception re-writing

        The hook is called while the cursor from the failed operation
        (if any) is still open and accessible.   Special cleanup operations
        can be called on this cursor; SQLAlchemy will attempt to close
        this cursor subsequent to this hook being invoked.

        As of SQLAlchemy 2.0, the "pre_ping" handler enabled using the
        :paramref:`_sa.create_engine.pool_pre_ping` parameter will also
        participate in the :meth:`.handle_error` process, **for those dialects
        that rely upon disconnect codes to detect database liveness**. Note
        that some dialects such as psycopg, psycopg2, and most MySQL dialects
        make use of a native ``ping()`` method supplied by the DBAPI which does
        not make use of disconnect codes.

        .. versionchanged:: 2.0.0 The :meth:`.DialectEvents.handle_error`
           event hook participates in connection pool "pre-ping" operations.
           Within this usage, the :attr:`.ExceptionContext.engine` attribute
           will be ``None``, however the :class:`.Dialect` in use is always
           available via the :attr:`.ExceptionContext.dialect` attribute.

        .. versionchanged:: 2.0.5 Added :attr:`.ExceptionContext.is_pre_ping`
           attribute which will be set to ``True`` when the
           :meth:`.DialectEvents.handle_error` event hook is triggered within
           a connection pool pre-ping operation.

        .. versionchanged:: 2.0.5 An issue was repaired that allows for the
           PostgreSQL ``psycopg`` and ``psycopg2`` drivers, as well as all
           MySQL drivers, to properly participate in the
           :meth:`.DialectEvents.handle_error` event hook during
           connection pool "pre-ping" operations; previously, the
           implementation was non-working for these drivers.


        A handler function has two options for replacing
        the SQLAlchemy-constructed exception into one that is user
        defined.   It can either raise this new exception directly, in
        which case all further event listeners are bypassed and the
        exception will be raised, after appropriate cleanup as taken
        place::

            @event.listens_for(Engine, "handle_error")
            def handle_exception(context):
                if isinstance(
                    context.original_exception, psycopg2.OperationalError
                ) and "failed" in str(context.original_exception):
                    raise MySpecialException("failed operation")

        .. warning::  Because the
           :meth:`_events.DialectEvents.handle_error`
           event specifically provides for exceptions to be re-thrown as
           the ultimate exception raised by the failed statement,
           **stack traces will be misleading** if the user-defined event
           handler itself fails and throws an unexpected exception;
           the stack trace may not illustrate the actual code line that
           failed!  It is advised to code carefully here and use
           logging and/or inline debugging if unexpected exceptions are
           occurring.

        Alternatively, a "chained" style of event handling can be
        used, by configuring the handler with the ``retval=True``
        modifier and returning the new exception instance from the
        function.  In this case, event handling will continue onto the
        next handler.   The "chained" exception is available using
        :attr:`.ExceptionContext.chained_exception`::

            @event.listens_for(Engine, "handle_error", retval=True)
            def handle_exception(context):
                if (
                    context.chained_exception is not None
                    and "special" in context.chained_exception.message
                ):
                    return MySpecialException(
                        "failed", cause=context.chained_exception
                    )

        Handlers that return ``None`` may be used within the chain; when
        a handler returns ``None``, the previous exception instance,
        if any, is maintained as the current exception that is passed onto the
        next handler.

        When a custom exception is raised or returned, SQLAlchemy raises
        this new exception as-is, it is not wrapped by any SQLAlchemy
        object.  If the exception is not a subclass of
        :class:`sqlalchemy.exc.StatementError`,
        certain features may not be available; currently this includes
        the ORM's feature of adding a detail hint about "autoflush" to
        exceptions raised within the autoflush process.

        :param context: an :class:`.ExceptionContext` object.  See this
         class for details on all available members.


        .. seealso::

            :ref:`pool_new_disconnect_codes`

        """

    def do_connect(
        self,
        dialect: Dialect,
        conn_rec: ConnectionPoolEntry,
        cargs: Tuple[Any, ...],
        cparams: Dict[str, Any],
    ) -> Optional[DBAPIConnection]:
        """Receive connection arguments before a connection is made.

        This event is useful in that it allows the handler to manipulate the
        cargs and/or cparams collections that control how the DBAPI
        ``connect()`` function will be called. ``cargs`` will always be a
        Python list that can be mutated in-place, and ``cparams`` a Python
        dictionary that may also be mutated::

            e = create_engine("postgresql+psycopg2://user@host/dbname")


            @event.listens_for(e, "do_connect")
            def receive_do_connect(dialect, conn_rec, cargs, cparams):
                cparams["password"] = "some_password"

        The event hook may also be used to override the call to ``connect()``
        entirely, by returning a non-``None`` DBAPI connection object::

            e = create_engine("postgresql+psycopg2://user@host/dbname")


            @event.listens_for(e, "do_connect")
            def receive_do_connect(dialect, conn_rec, cargs, cparams):
                return psycopg2.connect(*cargs, **cparams)

        .. seealso::

            :ref:`custom_dbapi_args`

        """

    def do_executemany(
        self,
        cursor: DBAPICursor,
        statement: str,
        parameters: _DBAPIMultiExecuteParams,
        context: ExecutionContext,
    ) -> Optional[Literal[True]]:
        """Receive a cursor to have executemany() called.

        Return the value True to halt further events from invoking,
        and to indicate that the cursor execution has already taken
        place within the event handler.

        """

    def do_execute_no_params(
        self, cursor: DBAPICursor, statement: str, context: ExecutionContext
    ) -> Optional[Literal[True]]:
        """Receive a cursor to have execute() with no parameters called.

        Return the value True to halt further events from invoking,
        and to indicate that the cursor execution has already taken
        place within the event handler.

        """

    def do_execute(
        self,
        cursor: DBAPICursor,
        statement: str,
        parameters: _DBAPISingleExecuteParams,
        context: ExecutionContext,
    ) -> Optional[Literal[True]]:
        """Receive a cursor to have execute() called.

        Return the value True to halt further events from invoking,
        and to indicate that the cursor execution has already taken
        place within the event handler.

        """

    def do_setinputsizes(
        self,
        inputsizes: Dict[BindParameter[Any], Any],
        cursor: DBAPICursor,
        statement: str,
        parameters: _DBAPIAnyExecuteParams,
        context: ExecutionContext,
    ) -> None:
        """Receive the setinputsizes dictionary for possible modification.

        This event is emitted in the case where the dialect makes use of the
        DBAPI ``cursor.setinputsizes()`` method which passes information about
        parameter binding for a particular statement.   The given
        ``inputsizes`` dictionary will contain :class:`.BindParameter` objects
        as keys, linked to DBAPI-specific type objects as values; for
        parameters that are not bound, they are added to the dictionary with
        ``None`` as the value, which means the parameter will not be included
        in the ultimate setinputsizes call.   The event may be used to inspect
        and/or log the datatypes that are being bound, as well as to modify the
        dictionary in place.  Parameters can be added, modified, or removed
        from this dictionary.   Callers will typically want to inspect the
        :attr:`.BindParameter.type` attribute of the given bind objects in
        order to make decisions about the DBAPI object.

        After the event, the ``inputsizes`` dictionary is converted into
        an appropriate datastructure to be passed to ``cursor.setinputsizes``;
        either a list for a positional bound parameter execution style,
        or a dictionary of string parameter keys to DBAPI type objects for
        a named bound parameter execution style.

        The setinputsizes hook overall is only used for dialects which include
        the flag ``use_setinputsizes=True``.  Dialects which use this
        include python-oracledb, cx_Oracle, pg8000, asyncpg, and pyodbc
        dialects.

        .. note::

            For use with pyodbc, the ``use_setinputsizes`` flag
            must be passed to the dialect, e.g.::

                create_engine("mssql+pyodbc://...", use_setinputsizes=True)

            .. seealso::

                  :ref:`mssql_pyodbc_setinputsizes`

        .. versionadded:: 1.2.9

        .. seealso::

            :ref:`cx_oracle_setinputsizes`

        """
        pass

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\window\expanding.py ===
from __future__ import annotations

from textwrap import dedent
from typing import (
    TYPE_CHECKING,
    Any,
    Callable,
    Literal,
)

from pandas.util._decorators import (
    deprecate_kwarg,
    doc,
)

from pandas.core.indexers.objects import (
    BaseIndexer,
    ExpandingIndexer,
    GroupbyIndexer,
)
from pandas.core.window.doc import (
    _shared_docs,
    create_section_header,
    kwargs_numeric_only,
    numba_notes,
    template_header,
    template_returns,
    template_see_also,
    window_agg_numba_parameters,
    window_apply_parameters,
)
from pandas.core.window.rolling import (
    BaseWindowGroupby,
    RollingAndExpandingMixin,
)

if TYPE_CHECKING:
    from pandas._typing import (
        Axis,
        QuantileInterpolation,
        WindowingRankType,
    )

    from pandas import (
        DataFrame,
        Series,
    )
    from pandas.core.generic import NDFrame


class Expanding(RollingAndExpandingMixin):
    """
    Provide expanding window calculations.

    Parameters
    ----------
    min_periods : int, default 1
        Minimum number of observations in window required to have a value;
        otherwise, result is ``np.nan``.

    axis : int or str, default 0
        If ``0`` or ``'index'``, roll across the rows.

        If ``1`` or ``'columns'``, roll across the columns.

        For `Series` this parameter is unused and defaults to 0.

    method : str {'single', 'table'}, default 'single'
        Execute the rolling operation per single column or row (``'single'``)
        or over the entire object (``'table'``).

        This argument is only implemented when specifying ``engine='numba'``
        in the method call.

        .. versionadded:: 1.3.0

    Returns
    -------
    pandas.api.typing.Expanding

    See Also
    --------
    rolling : Provides rolling window calculations.
    ewm : Provides exponential weighted functions.

    Notes
    -----
    See :ref:`Windowing Operations <window.expanding>` for further usage details
    and examples.

    Examples
    --------
    >>> df = pd.DataFrame({"B": [0, 1, 2, np.nan, 4]})
    >>> df
         B
    0  0.0
    1  1.0
    2  2.0
    3  NaN
    4  4.0

    **min_periods**

    Expanding sum with 1 vs 3 observations needed to calculate a value.

    >>> df.expanding(1).sum()
         B
    0  0.0
    1  1.0
    2  3.0
    3  3.0
    4  7.0
    >>> df.expanding(3).sum()
         B
    0  NaN
    1  NaN
    2  3.0
    3  3.0
    4  7.0
    """

    _attributes: list[str] = ["min_periods", "axis", "method"]

    def __init__(
        self,
        obj: NDFrame,
        min_periods: int = 1,
        axis: Axis = 0,
        method: str = "single",
        selection=None,
    ) -> None:
        super().__init__(
            obj=obj,
            min_periods=min_periods,
            axis=axis,
            method=method,
            selection=selection,
        )

    def _get_window_indexer(self) -> BaseIndexer:
        """
        Return an indexer class that will compute the window start and end bounds
        """
        return ExpandingIndexer()

    @doc(
        _shared_docs["aggregate"],
        see_also=dedent(
            """
        See Also
        --------
        pandas.DataFrame.aggregate : Similar DataFrame method.
        pandas.Series.aggregate : Similar Series method.
        """
        ),
        examples=dedent(
            """
        Examples
        --------
        >>> df = pd.DataFrame({"A": [1, 2, 3], "B": [4, 5, 6], "C": [7, 8, 9]})
        >>> df
           A  B  C
        0  1  4  7
        1  2  5  8
        2  3  6  9

        >>> df.ewm(alpha=0.5).mean()
                  A         B         C
        0  1.000000  4.000000  7.000000
        1  1.666667  4.666667  7.666667
        2  2.428571  5.428571  8.428571
        """
        ),
        klass="Series/Dataframe",
        axis="",
    )
    def aggregate(self, func, *args, **kwargs):
        return super().aggregate(func, *args, **kwargs)

    agg = aggregate

    @doc(
        template_header,
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser = pd.Series([1, 2, 3, 4], index=['a', 'b', 'c', 'd'])
        >>> ser.expanding().count()
        a    1.0
        b    2.0
        c    3.0
        d    4.0
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="count of non NaN observations",
        agg_method="count",
    )
    def count(self, numeric_only: bool = False):
        return super().count(numeric_only=numeric_only)

    @doc(
        template_header,
        create_section_header("Parameters"),
        window_apply_parameters,
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser = pd.Series([1, 2, 3, 4], index=['a', 'b', 'c', 'd'])
        >>> ser.expanding().apply(lambda s: s.max() - 2 * s.min())
        a   -1.0
        b    0.0
        c    1.0
        d    2.0
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="custom aggregation function",
        agg_method="apply",
    )
    def apply(
        self,
        func: Callable[..., Any],
        raw: bool = False,
        engine: Literal["cython", "numba"] | None = None,
        engine_kwargs: dict[str, bool] | None = None,
        args: tuple[Any, ...] | None = None,
        kwargs: dict[str, Any] | None = None,
    ):
        return super().apply(
            func,
            raw=raw,
            engine=engine,
            engine_kwargs=engine_kwargs,
            args=args,
            kwargs=kwargs,
        )

    @doc(
        template_header,
        create_section_header("Parameters"),
        kwargs_numeric_only,
        window_agg_numba_parameters(),
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Notes"),
        numba_notes,
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser = pd.Series([1, 2, 3, 4], index=['a', 'b', 'c', 'd'])
        >>> ser.expanding().sum()
        a     1.0
        b     3.0
        c     6.0
        d    10.0
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="sum",
        agg_method="sum",
    )
    def sum(
        self,
        numeric_only: bool = False,
        engine: Literal["cython", "numba"] | None = None,
        engine_kwargs: dict[str, bool] | None = None,
    ):
        return super().sum(
            numeric_only=numeric_only,
            engine=engine,
            engine_kwargs=engine_kwargs,
        )

    @doc(
        template_header,
        create_section_header("Parameters"),
        kwargs_numeric_only,
        window_agg_numba_parameters(),
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Notes"),
        numba_notes,
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser = pd.Series([3, 2, 1, 4], index=['a', 'b', 'c', 'd'])
        >>> ser.expanding().max()
        a    3.0
        b    3.0
        c    3.0
        d    4.0
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="maximum",
        agg_method="max",
    )
    def max(
        self,
        numeric_only: bool = False,
        engine: Literal["cython", "numba"] | None = None,
        engine_kwargs: dict[str, bool] | None = None,
    ):
        return super().max(
            numeric_only=numeric_only,
            engine=engine,
            engine_kwargs=engine_kwargs,
        )

    @doc(
        template_header,
        create_section_header("Parameters"),
        kwargs_numeric_only,
        window_agg_numba_parameters(),
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Notes"),
        numba_notes,
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser = pd.Series([2, 3, 4, 1], index=['a', 'b', 'c', 'd'])
        >>> ser.expanding().min()
        a    2.0
        b    2.0
        c    2.0
        d    1.0
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="minimum",
        agg_method="min",
    )
    def min(
        self,
        numeric_only: bool = False,
        engine: Literal["cython", "numba"] | None = None,
        engine_kwargs: dict[str, bool] | None = None,
    ):
        return super().min(
            numeric_only=numeric_only,
            engine=engine,
            engine_kwargs=engine_kwargs,
        )

    @doc(
        template_header,
        create_section_header("Parameters"),
        kwargs_numeric_only,
        window_agg_numba_parameters(),
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Notes"),
        numba_notes,
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser = pd.Series([1, 2, 3, 4], index=['a', 'b', 'c', 'd'])
        >>> ser.expanding().mean()
        a    1.0
        b    1.5
        c    2.0
        d    2.5
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="mean",
        agg_method="mean",
    )
    def mean(
        self,
        numeric_only: bool = False,
        engine: Literal["cython", "numba"] | None = None,
        engine_kwargs: dict[str, bool] | None = None,
    ):
        return super().mean(
            numeric_only=numeric_only,
            engine=engine,
            engine_kwargs=engine_kwargs,
        )

    @doc(
        template_header,
        create_section_header("Parameters"),
        kwargs_numeric_only,
        window_agg_numba_parameters(),
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Notes"),
        numba_notes,
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser = pd.Series([1, 2, 3, 4], index=['a', 'b', 'c', 'd'])
        >>> ser.expanding().median()
        a    1.0
        b    1.5
        c    2.0
        d    2.5
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="median",
        agg_method="median",
    )
    def median(
        self,
        numeric_only: bool = False,
        engine: Literal["cython", "numba"] | None = None,
        engine_kwargs: dict[str, bool] | None = None,
    ):
        return super().median(
            numeric_only=numeric_only,
            engine=engine,
            engine_kwargs=engine_kwargs,
        )

    @doc(
        template_header,
        create_section_header("Parameters"),
        dedent(
            """
        ddof : int, default 1
            Delta Degrees of Freedom.  The divisor used in calculations
            is ``N - ddof``, where ``N`` represents the number of elements.\n
        """
        ).replace("\n", "", 1),
        kwargs_numeric_only,
        window_agg_numba_parameters("1.4"),
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        "numpy.std : Equivalent method for NumPy array.\n",
        template_see_also,
        create_section_header("Notes"),
        dedent(
            """
        The default ``ddof`` of 1 used in :meth:`Series.std` is different
        than the default ``ddof`` of 0 in :func:`numpy.std`.

        A minimum of one period is required for the rolling calculation.\n
        """
        ).replace("\n", "", 1),
        create_section_header("Examples"),
        dedent(
            """
        >>> s = pd.Series([5, 5, 6, 7, 5, 5, 5])

        >>> s.expanding(3).std()
        0         NaN
        1         NaN
        2    0.577350
        3    0.957427
        4    0.894427
        5    0.836660
        6    0.786796
        dtype: float64
        """
        ).replace("\n", "", 1),
        window_method="expanding",
        aggregation_description="standard deviation",
        agg_method="std",
    )
    def std(
        self,
        ddof: int = 1,
        numeric_only: bool = False,
        engine: Literal["cython", "numba"] | None = None,
        engine_kwargs: dict[str, bool] | None = None,
    ):
        return super().std(
            ddof=ddof,
            numeric_only=numeric_only,
            engine=engine,
            engine_kwargs=engine_kwargs,
        )

    @doc(
        template_header,
        create_section_header("Parameters"),
        dedent(
            """
        ddof : int, default 1
            Delta Degrees of Freedom.  The divisor used in calculations
            is ``N - ddof``, where ``N`` represents the number of elements.\n
        """
        ).replace("\n", "", 1),
        kwargs_numeric_only,
        window_agg_numba_parameters("1.4"),
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        "numpy.var : Equivalent method for NumPy array.\n",
        template_see_also,
        create_section_header("Notes"),
        dedent(
            """
        The default ``ddof`` of 1 used in :meth:`Series.var` is different
        than the default ``ddof`` of 0 in :func:`numpy.var`.

        A minimum of one period is required for the rolling calculation.\n
        """
        ).replace("\n", "", 1),
        create_section_header("Examples"),
        dedent(
            """
        >>> s = pd.Series([5, 5, 6, 7, 5, 5, 5])

        >>> s.expanding(3).var()
        0         NaN
        1         NaN
        2    0.333333
        3    0.916667
        4    0.800000
        5    0.700000
        6    0.619048
        dtype: float64
        """
        ).replace("\n", "", 1),
        window_method="expanding",
        aggregation_description="variance",
        agg_method="var",
    )
    def var(
        self,
        ddof: int = 1,
        numeric_only: bool = False,
        engine: Literal["cython", "numba"] | None = None,
        engine_kwargs: dict[str, bool] | None = None,
    ):
        return super().var(
            ddof=ddof,
            numeric_only=numeric_only,
            engine=engine,
            engine_kwargs=engine_kwargs,
        )

    @doc(
        template_header,
        create_section_header("Parameters"),
        dedent(
            """
        ddof : int, default 1
            Delta Degrees of Freedom.  The divisor used in calculations
            is ``N - ddof``, where ``N`` represents the number of elements.\n
        """
        ).replace("\n", "", 1),
        kwargs_numeric_only,
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Notes"),
        "A minimum of one period is required for the calculation.\n\n",
        create_section_header("Examples"),
        dedent(
            """
        >>> s = pd.Series([0, 1, 2, 3])

        >>> s.expanding().sem()
        0         NaN
        1    0.707107
        2    0.707107
        3    0.745356
        dtype: float64
        """
        ).replace("\n", "", 1),
        window_method="expanding",
        aggregation_description="standard error of mean",
        agg_method="sem",
    )
    def sem(self, ddof: int = 1, numeric_only: bool = False):
        return super().sem(ddof=ddof, numeric_only=numeric_only)

    @doc(
        template_header,
        create_section_header("Parameters"),
        kwargs_numeric_only,
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        "scipy.stats.skew : Third moment of a probability density.\n",
        template_see_also,
        create_section_header("Notes"),
        "A minimum of three periods is required for the rolling calculation.\n\n",
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser = pd.Series([-1, 0, 2, -1, 2], index=['a', 'b', 'c', 'd', 'e'])
        >>> ser.expanding().skew()
        a         NaN
        b         NaN
        c    0.935220
        d    1.414214
        e    0.315356
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="unbiased skewness",
        agg_method="skew",
    )
    def skew(self, numeric_only: bool = False):
        return super().skew(numeric_only=numeric_only)

    @doc(
        template_header,
        create_section_header("Parameters"),
        kwargs_numeric_only,
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        "scipy.stats.kurtosis : Reference SciPy method.\n",
        template_see_also,
        create_section_header("Notes"),
        "A minimum of four periods is required for the calculation.\n\n",
        create_section_header("Examples"),
        dedent(
            """
        The example below will show a rolling calculation with a window size of
        four matching the equivalent function call using `scipy.stats`.

        >>> arr = [1, 2, 3, 4, 999]
        >>> import scipy.stats
        >>> print(f"{{scipy.stats.kurtosis(arr[:-1], bias=False):.6f}}")
        -1.200000
        >>> print(f"{{scipy.stats.kurtosis(arr, bias=False):.6f}}")
        4.999874
        >>> s = pd.Series(arr)
        >>> s.expanding(4).kurt()
        0         NaN
        1         NaN
        2         NaN
        3   -1.200000
        4    4.999874
        dtype: float64
        """
        ).replace("\n", "", 1),
        window_method="expanding",
        aggregation_description="Fisher's definition of kurtosis without bias",
        agg_method="kurt",
    )
    def kurt(self, numeric_only: bool = False):
        return super().kurt(numeric_only=numeric_only)

    @doc(
        template_header,
        create_section_header("Parameters"),
        dedent(
            """
        quantile : float
            Quantile to compute. 0 <= quantile <= 1.

            .. deprecated:: 2.1.0
                This will be renamed to 'q' in a future version.
        interpolation : {{'linear', 'lower', 'higher', 'midpoint', 'nearest'}}
            This optional parameter specifies the interpolation method to use,
            when the desired quantile lies between two data points `i` and `j`:

                * linear: `i + (j - i) * fraction`, where `fraction` is the
                  fractional part of the index surrounded by `i` and `j`.
                * lower: `i`.
                * higher: `j`.
                * nearest: `i` or `j` whichever is nearest.
                * midpoint: (`i` + `j`) / 2.
        """
        ).replace("\n", "", 1),
        kwargs_numeric_only,
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser = pd.Series([1, 2, 3, 4, 5, 6], index=['a', 'b', 'c', 'd', 'e', 'f'])
        >>> ser.expanding(min_periods=4).quantile(.25)
        a     NaN
        b     NaN
        c     NaN
        d    1.75
        e    2.00
        f    2.25
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="quantile",
        agg_method="quantile",
    )
    @deprecate_kwarg(old_arg_name="quantile", new_arg_name="q")
    def quantile(
        self,
        q: float,
        interpolation: QuantileInterpolation = "linear",
        numeric_only: bool = False,
    ):
        return super().quantile(
            q=q,
            interpolation=interpolation,
            numeric_only=numeric_only,
        )

    @doc(
        template_header,
        ".. versionadded:: 1.4.0 \n\n",
        create_section_header("Parameters"),
        dedent(
            """
        method : {{'average', 'min', 'max'}}, default 'average'
            How to rank the group of records that have the same value (i.e. ties):

            * average: average rank of the group
            * min: lowest rank in the group
            * max: highest rank in the group

        ascending : bool, default True
            Whether or not the elements should be ranked in ascending order.
        pct : bool, default False
            Whether or not to display the returned rankings in percentile
            form.
        """
        ).replace("\n", "", 1),
        kwargs_numeric_only,
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Examples"),
        dedent(
            """
        >>> s = pd.Series([1, 4, 2, 3, 5, 3])
        >>> s.expanding().rank()
        0    1.0
        1    2.0
        2    2.0
        3    3.0
        4    5.0
        5    3.5
        dtype: float64

        >>> s.expanding().rank(method="max")
        0    1.0
        1    2.0
        2    2.0
        3    3.0
        4    5.0
        5    4.0
        dtype: float64

        >>> s.expanding().rank(method="min")
        0    1.0
        1    2.0
        2    2.0
        3    3.0
        4    5.0
        5    3.0
        dtype: float64
        """
        ).replace("\n", "", 1),
        window_method="expanding",
        aggregation_description="rank",
        agg_method="rank",
    )
    def rank(
        self,
        method: WindowingRankType = "average",
        ascending: bool = True,
        pct: bool = False,
        numeric_only: bool = False,
    ):
        return super().rank(
            method=method,
            ascending=ascending,
            pct=pct,
            numeric_only=numeric_only,
        )

    @doc(
        template_header,
        create_section_header("Parameters"),
        dedent(
            """
        other : Series or DataFrame, optional
            If not supplied then will default to self and produce pairwise
            output.
        pairwise : bool, default None
            If False then only matching columns between self and other will be
            used and the output will be a DataFrame.
            If True then all pairwise combinations will be calculated and the
            output will be a MultiIndexed DataFrame in the case of DataFrame
            inputs. In the case of missing elements, only complete pairwise
            observations will be used.
        ddof : int, default 1
            Delta Degrees of Freedom.  The divisor used in calculations
            is ``N - ddof``, where ``N`` represents the number of elements.
        """
        ).replace("\n", "", 1),
        kwargs_numeric_only,
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        template_see_also,
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser1 = pd.Series([1, 2, 3, 4], index=['a', 'b', 'c', 'd'])
        >>> ser2 = pd.Series([10, 11, 13, 16], index=['a', 'b', 'c', 'd'])
        >>> ser1.expanding().cov(ser2)
        a         NaN
        b    0.500000
        c    1.500000
        d    3.333333
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="sample covariance",
        agg_method="cov",
    )
    def cov(
        self,
        other: DataFrame | Series | None = None,
        pairwise: bool | None = None,
        ddof: int = 1,
        numeric_only: bool = False,
    ):
        return super().cov(
            other=other,
            pairwise=pairwise,
            ddof=ddof,
            numeric_only=numeric_only,
        )

    @doc(
        template_header,
        create_section_header("Parameters"),
        dedent(
            """
        other : Series or DataFrame, optional
            If not supplied then will default to self and produce pairwise
            output.
        pairwise : bool, default None
            If False then only matching columns between self and other will be
            used and the output will be a DataFrame.
            If True then all pairwise combinations will be calculated and the
            output will be a MultiIndexed DataFrame in the case of DataFrame
            inputs. In the case of missing elements, only complete pairwise
            observations will be used.
        """
        ).replace("\n", "", 1),
        kwargs_numeric_only,
        create_section_header("Returns"),
        template_returns,
        create_section_header("See Also"),
        dedent(
            """
        cov : Similar method to calculate covariance.
        numpy.corrcoef : NumPy Pearson's correlation calculation.
        """
        ).replace("\n", "", 1),
        template_see_also,
        create_section_header("Notes"),
        dedent(
            """
        This function uses Pearson's definition of correlation
        (https://en.wikipedia.org/wiki/Pearson_correlation_coefficient).

        When `other` is not specified, the output will be self correlation (e.g.
        all 1's), except for :class:`~pandas.DataFrame` inputs with `pairwise`
        set to `True`.

        Function will return ``NaN`` for correlations of equal valued sequences;
        this is the result of a 0/0 division error.

        When `pairwise` is set to `False`, only matching columns between `self` and
        `other` will be used.

        When `pairwise` is set to `True`, the output will be a MultiIndex DataFrame
        with the original index on the first level, and the `other` DataFrame
        columns on the second level.

        In the case of missing elements, only complete pairwise observations
        will be used.\n
        """
        ),
        create_section_header("Examples"),
        dedent(
            """\
        >>> ser1 = pd.Series([1, 2, 3, 4], index=['a', 'b', 'c', 'd'])
        >>> ser2 = pd.Series([10, 11, 13, 16], index=['a', 'b', 'c', 'd'])
        >>> ser1.expanding().corr(ser2)
        a         NaN
        b    1.000000
        c    0.981981
        d    0.975900
        dtype: float64
        """
        ),
        window_method="expanding",
        aggregation_description="correlation",
        agg_method="corr",
    )
    def corr(
        self,
        other: DataFrame | Series | None = None,
        pairwise: bool | None = None,
        ddof: int = 1,
        numeric_only: bool = False,
    ):
        return super().corr(
            other=other,
            pairwise=pairwise,
            ddof=ddof,
            numeric_only=numeric_only,
        )


class ExpandingGroupby(BaseWindowGroupby, Expanding):
    """
    Provide a expanding groupby implementation.
    """

    _attributes = Expanding._attributes + BaseWindowGroupby._attributes

    def _get_window_indexer(self) -> GroupbyIndexer:
        """
        Return an indexer class that will compute the window start and end bounds

        Returns
        -------
        GroupbyIndexer
        """
        window_indexer = GroupbyIndexer(
            groupby_indices=self._grouper.indices,
            window_indexer=ExpandingIndexer,
        )
        return window_indexer

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_ssl.py ===
from __future__ import annotations

import contextlib
import operator as _operator
import ssl as _stdlib_ssl
from enum import Enum as _Enum
from typing import TYPE_CHECKING, Any, ClassVar, Final as TFinal, Generic, TypeVar

import trio

from . import _sync
from ._highlevel_generic import aclose_forcefully
from ._util import ConflictDetector, final
from .abc import Listener, Stream

if TYPE_CHECKING:
    from collections.abc import Awaitable, Callable

    from typing_extensions import TypeVarTuple, Unpack

    Ts = TypeVarTuple("Ts")

# General theory of operation:
#
# We implement an API that closely mirrors the stdlib ssl module's blocking
# API, and we do it using the stdlib ssl module's non-blocking in-memory API.
# The stdlib non-blocking in-memory API is barely documented, and acts as a
# thin wrapper around openssl, whose documentation also leaves something to be
# desired. So here's the main things you need to know to understand the code
# in this file:
#
# We use an ssl.SSLObject, which exposes the four main I/O operations:
#
# - do_handshake: performs the initial handshake. Must be called once at the
#   beginning of each connection; is a no-op once it's completed once.
#
# - write: takes some unencrypted data and attempts to send it to the remote
#   peer.

# - read: attempts to decrypt and return some data from the remote peer.
#
# - unwrap: this is weirdly named; maybe it helps to realize that the thing it
#   wraps is called SSL_shutdown. It sends a cryptographically signed message
#   saying "I'm closing this connection now", and then waits to receive the
#   same from the remote peer (unless we already received one, in which case
#   it returns immediately).
#
# All of these operations read and write from some in-memory buffers called
# "BIOs", which are an opaque OpenSSL-specific object that's basically
# semantically equivalent to a Python bytearray. When they want to send some
# bytes to the remote peer, they append them to the outgoing BIO, and when
# they want to receive some bytes from the remote peer, they try to pull them
# out of the incoming BIO. "Sending" always succeeds, because the outgoing BIO
# can always be extended to hold more data. "Receiving" acts sort of like a
# non-blocking socket: it might manage to get some data immediately, or it
# might fail and need to be tried again later. We can also directly add or
# remove data from the BIOs whenever we want.
#
# Now the problem is that while these I/O operations are opaque atomic
# operations from the point of view of us calling them, under the hood they
# might require some arbitrary sequence of sends and receives from the remote
# peer. This is particularly true for do_handshake, which generally requires a
# few round trips, but it's also true for write and read, due to an evil thing
# called "renegotiation".
#
# Renegotiation is the process by which one of the peers might arbitrarily
# decide to redo the handshake at any time. Did I mention it's evil? It's
# pretty evil, and almost universally hated. The HTTP/2 spec forbids the use
# of TLS renegotiation for HTTP/2 connections. TLS 1.3 removes it from the
# protocol entirely. It's impossible to trigger a renegotiation if using
# Python's ssl module. OpenSSL's renegotiation support is pretty buggy [1].
# Nonetheless, it does get used in real life, mostly in two cases:
#
# 1) Normally in TLS 1.2 and below, when the client side of a connection wants
# to present a certificate to prove their identity, that certificate gets sent
# in plaintext. This is bad, because it means that anyone eavesdropping can
# see who's connecting – it's like sending your username in plain text. Not as
# bad as sending your password in plain text, but still, pretty bad. However,
# renegotiations *are* encrypted. So as a workaround, it's not uncommon for
# systems that want to use client certificates to first do an anonymous
# handshake, and then to turn around and do a second handshake (=
# renegotiation) and this time ask for a client cert. Or sometimes this is
# done on a case-by-case basis, e.g. a web server might accept a connection,
# read the request, and then once it sees the page you're asking for it might
# stop and ask you for a certificate.
#
# 2) In principle the same TLS connection can be used for an arbitrarily long
# time, and might transmit arbitrarily large amounts of data. But this creates
# a cryptographic problem: an attacker who has access to arbitrarily large
# amounts of data that's all encrypted using the same key may eventually be
# able to use this to figure out the key. Is this a real practical problem? I
# have no idea, I'm not a cryptographer. In any case, some people worry that
# it's a problem, so their TLS libraries are designed to automatically trigger
# a renegotiation every once in a while on some sort of timer.
#
# The end result is that you might be going along, minding your own business,
# and then *bam*! a wild renegotiation appears! And you just have to cope.
#
# The reason that coping with renegotiations is difficult is that some
# unassuming "read" or "write" call might find itself unable to progress until
# it does a handshake, which remember is a process with multiple round
# trips. So read might have to send data, and write might have to receive
# data, and this might happen multiple times. And some of those attempts might
# fail because there isn't any data yet, and need to be retried. Managing all
# this is pretty complicated.
#
# Here's how openssl (and thus the stdlib ssl module) handle this. All of the
# I/O operations above follow the same rules. When you call one of them:
#
# - it might write some data to the outgoing BIO
# - it might read some data from the incoming BIO
# - it might raise SSLWantReadError if it can't complete without reading more
#   data from the incoming BIO. This is important: the "read" in ReadError
#   refers to reading from the *underlying* stream.
# - (and in principle it might raise SSLWantWriteError too, but that never
#   happens when using memory BIOs, so never mind)
#
# If it doesn't raise an error, then the operation completed successfully
# (though we still need to take any outgoing data out of the memory buffer and
# put it onto the wire). If it *does* raise an error, then we need to retry
# *exactly that method call* later – in particular, if a 'write' failed, we
# need to try again later *with the same data*, because openssl might have
# already committed some of the initial parts of our data to its output even
# though it didn't tell us that, and has remembered that the next time we call
# write it needs to skip the first 1024 bytes or whatever it is. (Well,
# technically, we're actually allowed to call 'write' again with a data buffer
# which is the same as our old one PLUS some extra stuff added onto the end,
# but in Trio that never comes up so never mind.)
#
# There are some people online who claim that once you've gotten a Want*Error
# then the *very next call* you make to openssl *must* be the same as the
# previous one. I'm pretty sure those people are wrong. In particular, it's
# okay to call write, get a WantReadError, and then call read a few times;
# it's just that *the next time you call write*, it has to be with the same
# data.
#
# One final wrinkle: we want our SSLStream to support full-duplex operation,
# i.e. it should be possible for one task to be calling send_all while another
# task is calling receive_some. But renegotiation makes this a big hassle, because
# even if SSLStream's restricts themselves to one task calling send_all and one
# task calling receive_some, those two tasks might end up both wanting to call
# send_all, or both to call receive_some at the same time *on the underlying
# stream*. So we have to do some careful locking to hide this problem from our
# users.
#
# (Renegotiation is evil.)
#
# So our basic strategy is to define a single helper method called "_retry",
# which has generic logic for dealing with SSLWantReadError, pushing data from
# the outgoing BIO to the wire, reading data from the wire to the incoming
# BIO, retrying an I/O call until it works, and synchronizing with other tasks
# that might be calling _retry concurrently. Basically it takes an SSLObject
# non-blocking in-memory method and converts it into a Trio async blocking
# method. _retry is only about 30 lines of code, but all these cases
# multiplied by concurrent calls make it extremely tricky, so there are lots
# of comments down below on the details, and a really extensive test suite in
# test_ssl.py. And now you know *why* it's so tricky, and can probably
# understand how it works.
#
# [1] https://rt.openssl.org/Ticket/Display.html?id=3712

# XX how closely should we match the stdlib API?
# - maybe suppress_ragged_eofs=False is a better default?
# - maybe check crypto folks for advice?
# - this is also interesting: https://bugs.python.org/issue8108#msg102867

# Definitely keep an eye on Cory's TLS API ideas on security-sig etc.

# XX document behavior on cancellation/error (i.e.: all is lost abandon
# stream)
# docs will need to make very clear that this is different from all the other
# cancellations in core Trio


T = TypeVar("T")

################################################################
# SSLStream
################################################################

# Ideally, when the user calls SSLStream.receive_some() with no argument, then
# we should do exactly one call to self.transport_stream.receive_some(),
# decrypt everything we got, and return it. Unfortunately, the way openssl's
# API works, we have to pick how much data we want to allow when we call
# read(), and then it (potentially) triggers a call to
# transport_stream.receive_some(). So at the time we pick the amount of data
# to decrypt, we don't know how much data we've read. As a simple heuristic,
# we record the max amount of data returned by previous calls to
# transport_stream.receive_some(), and we use that for future calls to read().
# But what do we use for the very first call? That's what this constant sets.
#
# Note that the value passed to read() is a limit on the amount of
# *decrypted* data, but we can only see the size of the *encrypted* data
# returned by transport_stream.receive_some(). TLS adds a small amount of
# framing overhead, and TLS compression is rarely used these days because it's
# insecure. So the size of the encrypted data should be a slight over-estimate
# of the size of the decrypted data, which is exactly what we want.
#
# The specific value is not really based on anything; it might be worth tuning
# at some point. But, if you have an TCP connection with the typical 1500 byte
# MTU and an initial window of 10 (see RFC 6928), then the initial burst of
# data will be limited to ~15000 bytes (or a bit less due to IP-level framing
# overhead), so this is chosen to be larger than that.
STARTING_RECEIVE_SIZE: TFinal = 16384


def _is_eof(exc: BaseException | None) -> bool:
    # There appears to be a bug on Python 3.10, where SSLErrors
    # aren't properly translated into SSLEOFErrors.
    # This stringly-typed error check is borrowed from the AnyIO
    # project.
    return isinstance(exc, _stdlib_ssl.SSLEOFError) or (
        "UNEXPECTED_EOF_WHILE_READING" in getattr(exc, "strerror", ())
    )


class NeedHandshakeError(Exception):
    """Some :class:`SSLStream` methods can't return any meaningful data until
    after the handshake. If you call them before the handshake, they raise
    this error.

    """


class _Once:
    __slots__ = ("_afn", "_args", "_done", "started")

    def __init__(
        self,
        afn: Callable[[*Ts], Awaitable[object]],
        *args: Unpack[Ts],
    ) -> None:
        self._afn = afn
        self._args = args
        self.started = False
        self._done = _sync.Event()

    async def ensure(self, *, checkpoint: bool) -> None:
        if not self.started:
            self.started = True
            await self._afn(*self._args)
            self._done.set()
        elif not checkpoint and self._done.is_set():
            return
        else:
            await self._done.wait()

    @property
    def done(self) -> bool:
        return bool(self._done.is_set())


_State = _Enum("_State", ["OK", "BROKEN", "CLOSED"])

# invariant
T_Stream = TypeVar("T_Stream", bound=Stream)


@final
class SSLStream(Stream, Generic[T_Stream]):
    r"""Encrypted communication using SSL/TLS.

    :class:`SSLStream` wraps an arbitrary :class:`~trio.abc.Stream`, and
    allows you to perform encrypted communication over it using the usual
    :class:`~trio.abc.Stream` interface. You pass regular data to
    :meth:`send_all`, then it encrypts it and sends the encrypted data on the
    underlying :class:`~trio.abc.Stream`; :meth:`receive_some` takes encrypted
    data out of the underlying :class:`~trio.abc.Stream` and decrypts it
    before returning it.

    You should read the standard library's :mod:`ssl` documentation carefully
    before attempting to use this class, and probably other general
    documentation on SSL/TLS as well. SSL/TLS is subtle and quick to
    anger. Really. I'm not kidding.

    Args:
      transport_stream (~trio.abc.Stream): The stream used to transport
          encrypted data. Required.

      ssl_context (~ssl.SSLContext): The :class:`~ssl.SSLContext` used for
          this connection. Required. Usually created by calling
          :func:`ssl.create_default_context`.

      server_hostname (str, bytes, or None): The name of the server being
          connected to. Used for `SNI
          <https://en.wikipedia.org/wiki/Server_Name_Indication>`__ and for
          validating the server's certificate (if hostname checking is
          enabled). This is effectively mandatory for clients, and actually
          mandatory if ``ssl_context.check_hostname`` is ``True``.

      server_side (bool): Whether this stream is acting as a client or
          server. Defaults to False, i.e. client mode.

      https_compatible (bool): There are two versions of SSL/TLS commonly
          encountered in the wild: the standard version, and the version used
          for HTTPS (HTTP-over-SSL/TLS).

          Standard-compliant SSL/TLS implementations always send a
          cryptographically signed ``close_notify`` message before closing the
          connection. This is important because if the underlying transport
          were simply closed, then there wouldn't be any way for the other
          side to know whether the connection was intentionally closed by the
          peer that they negotiated a cryptographic connection to, or by some
          `man-in-the-middle
          <https://en.wikipedia.org/wiki/Man-in-the-middle_attack>`__ attacker
          who can't manipulate the cryptographic stream, but can manipulate
          the transport layer (a so-called "truncation attack").

          However, this part of the standard is widely ignored by real-world
          HTTPS implementations, which means that if you want to interoperate
          with them, then you NEED to ignore it too.

          Fortunately this isn't as bad as it sounds, because the HTTP
          protocol already includes its own equivalent of ``close_notify``, so
          doing this again at the SSL/TLS level is redundant. But not all
          protocols do! Therefore, by default Trio implements the safer
          standard-compliant version (``https_compatible=False``). But if
          you're speaking HTTPS or some other protocol where
          ``close_notify``\s are commonly skipped, then you should set
          ``https_compatible=True``; with this setting, Trio will neither
          expect nor send ``close_notify`` messages.

          If you have code that was written to use :class:`ssl.SSLSocket` and
          now you're porting it to Trio, then it may be useful to know that a
          difference between :class:`SSLStream` and :class:`ssl.SSLSocket` is
          that :class:`~ssl.SSLSocket` implements the
          ``https_compatible=True`` behavior by default.

    Attributes:
      transport_stream (trio.abc.Stream): The underlying transport stream
          that was passed to ``__init__``. An example of when this would be
          useful is if you're using :class:`SSLStream` over a
          :class:`~trio.SocketStream` and want to call the
          :class:`~trio.SocketStream`'s :meth:`~trio.SocketStream.setsockopt`
          method.

    Internally, this class is implemented using an instance of
    :class:`ssl.SSLObject`, and all of :class:`~ssl.SSLObject`'s methods and
    attributes are re-exported as methods and attributes on this class.
    However, there is one difference: :class:`~ssl.SSLObject` has several
    methods that return information about the encrypted connection, like
    :meth:`~ssl.SSLSocket.cipher` or
    :meth:`~ssl.SSLSocket.selected_alpn_protocol`. If you call them before the
    handshake, when they can't possibly return useful data, then
    :class:`ssl.SSLObject` returns None, but :class:`trio.SSLStream`
    raises :exc:`NeedHandshakeError`.

    This also means that if you register a SNI callback using
    `~ssl.SSLContext.sni_callback`, then the first argument your callback
    receives will be a :class:`ssl.SSLObject`.

    """

    # Note: any new arguments here should likely also be added to
    # SSLListener.__init__, and maybe the open_ssl_over_tcp_* helpers.
    def __init__(
        self,
        transport_stream: T_Stream,
        ssl_context: _stdlib_ssl.SSLContext,
        *,
        server_hostname: str | bytes | None = None,
        server_side: bool = False,
        https_compatible: bool = False,
    ) -> None:
        self.transport_stream: T_Stream = transport_stream
        self._state = _State.OK
        self._https_compatible = https_compatible
        self._outgoing = _stdlib_ssl.MemoryBIO()
        self._delayed_outgoing: bytes | None = None
        self._incoming = _stdlib_ssl.MemoryBIO()
        self._ssl_object = ssl_context.wrap_bio(
            self._incoming,
            self._outgoing,
            server_side=server_side,
            server_hostname=server_hostname,
        )
        # Tracks whether we've already done the initial handshake
        self._handshook = _Once(self._do_handshake)

        # These are used to synchronize access to self.transport_stream
        self._inner_send_lock = _sync.StrictFIFOLock()
        self._inner_recv_count = 0
        self._inner_recv_lock = _sync.Lock()

        # These are used to make sure that our caller doesn't attempt to make
        # multiple concurrent calls to send_all/wait_send_all_might_not_block
        # or to receive_some.
        self._outer_send_conflict_detector = ConflictDetector(
            "another task is currently sending data on this SSLStream",
        )
        self._outer_recv_conflict_detector = ConflictDetector(
            "another task is currently receiving data on this SSLStream",
        )

        self._estimated_receive_size = STARTING_RECEIVE_SIZE

    _forwarded: ClassVar = {
        "context",
        "server_side",
        "server_hostname",
        "session",
        "session_reused",
        "getpeercert",
        "selected_npn_protocol",
        "cipher",
        "shared_ciphers",
        "compression",
        "pending",
        "get_channel_binding",
        "selected_alpn_protocol",
        "version",
    }

    _after_handshake: ClassVar = {
        "session_reused",
        "getpeercert",
        "selected_npn_protocol",
        "cipher",
        "shared_ciphers",
        "compression",
        "get_channel_binding",
        "selected_alpn_protocol",
        "version",
    }

    def __getattr__(  # type: ignore[explicit-any]
        self,
        name: str,
    ) -> Any:
        if name in self._forwarded:
            if name in self._after_handshake and not self._handshook.done:
                raise NeedHandshakeError(f"call do_handshake() before calling {name!r}")

            return getattr(self._ssl_object, name)
        else:
            raise AttributeError(name)

    def __setattr__(self, name: str, value: object) -> None:
        if name in self._forwarded:
            setattr(self._ssl_object, name, value)
        else:
            super().__setattr__(name, value)

    def __dir__(self) -> list[str]:
        return list(super().__dir__()) + list(self._forwarded)

    def _check_status(self) -> None:
        if self._state is _State.OK:
            return
        elif self._state is _State.BROKEN:
            raise trio.BrokenResourceError
        elif self._state is _State.CLOSED:
            raise trio.ClosedResourceError
        else:  # pragma: no cover
            raise AssertionError()

    # This is probably the single trickiest function in Trio. It has lots of
    # comments, though, just make sure to think carefully if you ever have to
    # touch it. The big comment at the top of this file will help explain
    # too.
    async def _retry(
        self,
        fn: Callable[[*Ts], T],
        *args: Unpack[Ts],
        ignore_want_read: bool = False,
        is_handshake: bool = False,
    ) -> T | None:
        await trio.lowlevel.checkpoint_if_cancelled()
        yielded = False
        finished = False
        while not finished:
            # WARNING: this code needs to be very careful with when it
            # calls 'await'! There might be multiple tasks calling this
            # function at the same time trying to do different operations,
            # so we need to be careful to:
            #
            # 1) interact with the SSLObject, then
            # 2) await on exactly one thing that lets us make forward
            # progress, then
            # 3) loop or exit
            #
            # In particular we don't want to yield while interacting with
            # the SSLObject (because it's shared state, so someone else
            # might come in and mess with it while we're suspended), and
            # we don't want to yield *before* starting the operation that
            # will help us make progress, because then someone else might
            # come in and leapfrog us.

            # Call the SSLObject method, and get its result.
            #
            # NB: despite what the docs say, SSLWantWriteError can't
            # happen – "Writes to memory BIOs will always succeed if
            # memory is available: that is their size can grow
            # indefinitely."
            # https://wiki.openssl.org/index.php/Manual:BIO_s_mem(3)
            want_read = False
            ret = None
            try:
                ret = fn(*args)
            except _stdlib_ssl.SSLWantReadError:
                want_read = True
            except (_stdlib_ssl.SSLError, _stdlib_ssl.CertificateError) as exc:
                self._state = _State.BROKEN
                raise trio.BrokenResourceError from exc
            else:
                finished = True
            if ignore_want_read:
                want_read = False
                finished = True
            to_send = self._outgoing.read()

            # Some versions of SSL_do_handshake have a bug in how they handle
            # the TLS 1.3 handshake on the server side: after the handshake
            # finishes, they automatically send session tickets, even though
            # the client may not be expecting data to arrive at this point and
            # sending it could cause a deadlock or lost data. This applies at
            # least to OpenSSL 1.1.1c and earlier, and the OpenSSL devs
            # currently have no plans to fix it:
            #
            #   https://github.com/openssl/openssl/issues/7948
            #   https://github.com/openssl/openssl/issues/7967
            #
            # The correct behavior is to wait to send session tickets on the
            # first call to SSL_write. (This is what BoringSSL does.) So, we
            # use a heuristic to detect when OpenSSL has tried to send session
            # tickets, and we manually delay sending them until the
            # appropriate moment. For more discussion see:
            #
            #   https://github.com/python-trio/trio/issues/819#issuecomment-517529763
            if (
                is_handshake
                and not want_read
                and self._ssl_object.server_side
                and self._ssl_object.version() == "TLSv1.3"
            ):
                assert self._delayed_outgoing is None
                self._delayed_outgoing = to_send
                to_send = b""

            # Outputs from the above code block are:
            #
            # - to_send: bytestring; if non-empty then we need to send
            #   this data to make forward progress
            #
            # - want_read: True if we need to receive_some some data to make
            #   forward progress
            #
            # - finished: False means that we need to retry the call to
            #   fn(*args) again, after having pushed things forward. True
            #   means we still need to do whatever was said (in particular
            #   send any data in to_send), but once we do then we're
            #   done.
            #
            # - ret: the operation's return value. (Meaningless unless
            #   finished is True.)
            #
            # Invariant: want_read and finished can't both be True at the
            # same time.
            #
            # Now we need to move things forward. There are two things we
            # might have to do, and any given operation might require
            # either, both, or neither to proceed:
            #
            # - send the data in to_send
            #
            # - receive_some some data and put it into the incoming BIO
            #
            # Our strategy is: if there's data to send, send it;
            # *otherwise* if there's data to receive_some, receive_some it.
            #
            # If both need to happen, then we only send. Why? Well, we
            # know that *right now* we have to both send and receive_some
            # before the operation can complete. But as soon as we yield,
            # that information becomes potentially stale – e.g. while
            # we're sending, some other task might go and receive_some the
            # data we need and put it into the incoming BIO. And if it
            # does, then we *definitely don't* want to do a receive_some –
            # there might not be any more data coming, and we'd deadlock!
            # We could do something tricky to keep track of whether a
            # receive_some happens while we're sending, but the case where
            # we have to do both is very unusual (only during a
            # renegotiation), so it's better to keep things simple. So we
            # do just one potentially-blocking operation, then check again
            # for fresh information.
            #
            # And we prioritize sending over receiving because, if there
            # are multiple tasks that want to receive_some, then it
            # doesn't matter what order they go in. But if there are
            # multiple tasks that want to send, then they each have
            # different data, and the data needs to get put onto the wire
            # in the same order that it was retrieved from the outgoing
            # BIO. So if we have data to send, that *needs* to be the
            # *very* *next* *thing* we do, to make sure no-one else sneaks
            # in before us. Or if we can't send immediately because
            # someone else is, then we at least need to get in line
            # immediately.
            if to_send:
                # NOTE: This relies on the lock being strict FIFO fair!
                async with self._inner_send_lock:
                    yielded = True
                    try:
                        if self._delayed_outgoing is not None:
                            to_send = self._delayed_outgoing + to_send
                            self._delayed_outgoing = None
                        await self.transport_stream.send_all(to_send)
                    except:
                        # Some unknown amount of our data got sent, and we
                        # don't know how much. This stream is doomed.
                        self._state = _State.BROKEN
                        raise
            elif want_read:
                # It's possible that someone else is already blocked in
                # transport_stream.receive_some. If so then we want to
                # wait for them to finish, but we don't want to call
                # transport_stream.receive_some again ourselves; we just
                # want to loop around and check if their contribution
                # helped anything. So we make a note of how many times
                # some task has been through here before taking the lock,
                # and if it's changed by the time we get the lock, then we
                # skip calling transport_stream.receive_some and loop
                # around immediately.
                recv_count = self._inner_recv_count
                async with self._inner_recv_lock:
                    yielded = True
                    if recv_count == self._inner_recv_count:
                        data = await self.transport_stream.receive_some()
                        if not data:
                            self._incoming.write_eof()
                        else:
                            self._estimated_receive_size = max(
                                self._estimated_receive_size,
                                len(data),
                            )
                            self._incoming.write(data)
                        self._inner_recv_count += 1
        if not yielded:
            await trio.lowlevel.cancel_shielded_checkpoint()
        return ret

    async def _do_handshake(self) -> None:
        try:
            await self._retry(self._ssl_object.do_handshake, is_handshake=True)
        except:
            self._state = _State.BROKEN
            raise

    async def do_handshake(self) -> None:
        """Ensure that the initial handshake has completed.

        The SSL protocol requires an initial handshake to exchange
        certificates, select cryptographic keys, and so forth, before any
        actual data can be sent or received. You don't have to call this
        method; if you don't, then :class:`SSLStream` will automatically
        perform the handshake as needed, the first time you try to send or
        receive data. But if you want to trigger it manually – for example,
        because you want to look at the peer's certificate before you start
        talking to them – then you can call this method.

        If the initial handshake is already in progress in another task, this
        waits for it to complete and then returns.

        If the initial handshake has already completed, this returns
        immediately without doing anything (except executing a checkpoint).

        .. warning:: If this method is cancelled, then it may leave the
           :class:`SSLStream` in an unusable state. If this happens then any
           future attempt to use the object will raise
           :exc:`trio.BrokenResourceError`.

        """
        self._check_status()
        await self._handshook.ensure(checkpoint=True)

    # Most things work if we don't explicitly force do_handshake to be called
    # before calling receive_some or send_all, because openssl will
    # automatically perform the handshake on the first SSL_{read,write}
    # call. BUT, allowing openssl to do this will disable Python's hostname
    # checking!!! See:
    #   https://bugs.python.org/issue30141
    # So we *definitely* have to make sure that do_handshake is called
    # before doing anything else.
    async def receive_some(self, max_bytes: int | None = None) -> bytes | bytearray:
        """Read some data from the underlying transport, decrypt it, and
        return it.

        See :meth:`trio.abc.ReceiveStream.receive_some` for details.

        .. warning:: If this method is cancelled while the initial handshake
           or a renegotiation are in progress, then it may leave the
           :class:`SSLStream` in an unusable state. If this happens then any
           future attempt to use the object will raise
           :exc:`trio.BrokenResourceError`.

        """
        with self._outer_recv_conflict_detector:
            self._check_status()
            try:
                await self._handshook.ensure(checkpoint=False)
            except trio.BrokenResourceError as exc:
                # For some reason, EOF before handshake sometimes raises
                # SSLSyscallError instead of SSLEOFError (e.g. on my linux
                # laptop, but not on appveyor). Thanks openssl.
                if self._https_compatible and (
                    isinstance(exc.__cause__, _stdlib_ssl.SSLSyscallError)
                    or _is_eof(exc.__cause__)
                ):
                    await trio.lowlevel.checkpoint()
                    return b""
                else:
                    raise
            if max_bytes is None:
                # If we somehow have more data already in our pending buffer
                # than the estimate receive size, bump up our size a bit for
                # this read only.
                max_bytes = max(self._estimated_receive_size, self._incoming.pending)
            else:
                max_bytes = _operator.index(max_bytes)
                if max_bytes < 1:
                    raise ValueError("max_bytes must be >= 1")
            try:
                received = await self._retry(self._ssl_object.read, max_bytes)
                assert received is not None
                return received
            except trio.BrokenResourceError as exc:
                # This isn't quite equivalent to just returning b"" in the
                # first place, because we still end up with self._state set to
                # BROKEN. But that's actually fine, because after getting an
                # EOF on TLS then the only thing you can do is close the
                # stream, and closing doesn't care about the state.

                if self._https_compatible and _is_eof(exc.__cause__):
                    await trio.lowlevel.checkpoint()
                    return b""
                else:
                    raise

    async def send_all(self, data: bytes | bytearray | memoryview) -> None:
        """Encrypt some data and then send it on the underlying transport.

        See :meth:`trio.abc.SendStream.send_all` for details.

        .. warning:: If this method is cancelled, then it may leave the
           :class:`SSLStream` in an unusable state. If this happens then any
           attempt to use the object will raise
           :exc:`trio.BrokenResourceError`.

        """
        with self._outer_send_conflict_detector:
            self._check_status()
            await self._handshook.ensure(checkpoint=False)
            # SSLObject interprets write(b"") as an EOF for some reason, which
            # is not what we want.
            if not data:
                await trio.lowlevel.checkpoint()
                return
            await self._retry(self._ssl_object.write, data)

    async def unwrap(self) -> tuple[Stream, bytes | bytearray]:
        """Cleanly close down the SSL/TLS encryption layer, allowing the
        underlying stream to be used for unencrypted communication.

        You almost certainly don't need this.

        Returns:
          A pair ``(transport_stream, trailing_bytes)``, where
          ``transport_stream`` is the underlying transport stream, and
          ``trailing_bytes`` is a byte string. Since :class:`SSLStream`
          doesn't necessarily know where the end of the encrypted data will
          be, it can happen that it accidentally reads too much from the
          underlying stream. ``trailing_bytes`` contains this extra data; you
          should process it as if it was returned from a call to
          ``transport_stream.receive_some(...)``.

        """
        with self._outer_recv_conflict_detector, self._outer_send_conflict_detector:
            self._check_status()
            await self._handshook.ensure(checkpoint=False)
            await self._retry(self._ssl_object.unwrap)
            transport_stream = self.transport_stream
            self._state = _State.CLOSED
            self.transport_stream = None  # type: ignore[assignment]  # State is CLOSED now, nothing should use
            return (transport_stream, self._incoming.read())

    async def aclose(self) -> None:
        """Gracefully shut down this connection, and close the underlying
        transport.

        If ``https_compatible`` is False (the default), then this attempts to
        first send a ``close_notify`` and then close the underlying stream by
        calling its :meth:`~trio.abc.AsyncResource.aclose` method.

        If ``https_compatible`` is set to True, then this simply closes the
        underlying stream and marks this stream as closed.

        """
        if self._state is _State.CLOSED:
            await trio.lowlevel.checkpoint()
            return
        if self._state is _State.BROKEN or self._https_compatible:
            self._state = _State.CLOSED
            await self.transport_stream.aclose()
            return
        try:
            # https_compatible=False, so we're in spec-compliant mode and have
            # to send close_notify so that the other side gets a cryptographic
            # assurance that we've called aclose. Of course, we can't do
            # anything cryptographic until after we've completed the
            # handshake:
            await self._handshook.ensure(checkpoint=False)
            # Then, we call SSL_shutdown *once*, because we want to send a
            # close_notify but *not* wait for the other side to send back a
            # response. In principle it would be more polite to wait for the
            # other side to reply with their own close_notify. However, if
            # they aren't paying attention (e.g., if they're just sending
            # data and not receiving) then we will never notice our
            # close_notify and we'll be waiting forever. Eventually we'll time
            # out (hopefully), but it's still kind of nasty. And we can't
            # require the other side to always be receiving, because (a)
            # backpressure is kind of important, and (b) I bet there are
            # broken TLS implementations out there that don't receive all the
            # time. (Like e.g. anyone using Python ssl in synchronous mode.)
            #
            # The send-then-immediately-close behavior is explicitly allowed
            # by the TLS specs, so we're ok on that.
            #
            # Subtlety: SSLObject.unwrap will immediately call it a second
            # time, and the second time will raise SSLWantReadError because
            # there hasn't been time for the other side to respond
            # yet. (Unless they spontaneously sent a close_notify before we
            # called this, and it's either already been processed or gets
            # pulled out of the buffer by Python's second call.) So the way to
            # do what we want is to ignore SSLWantReadError on this call.
            #
            # Also, because the other side might have already sent
            # close_notify and closed their connection then it's possible that
            # our attempt to send close_notify will raise
            # BrokenResourceError. This is totally legal, and in fact can happen
            # with two well-behaved Trio programs talking to each other, so we
            # don't want to raise an error. So we suppress BrokenResourceError
            # here. (This is safe, because literally the only thing this call
            # to _retry will do is send the close_notify alert, so that's
            # surely where the error comes from.)
            #
            # FYI in some cases this could also raise SSLSyscallError which I
            # think is because SSL_shutdown is terrible. (Check out that note
            # at the bottom of the man page saying that it sometimes gets
            # raised spuriously.) I haven't seen this since we switched to
            # immediately closing the socket, and I don't know exactly what
            # conditions cause it and how to respond, so for now we're just
            # letting that happen. But if you start seeing it, then hopefully
            # this will give you a little head start on tracking it down,
            # because whoa did this puzzle us at the 2017 PyCon sprints.
            #
            # Also, if someone else is blocked in send/receive, then we aren't
            # going to be able to do a clean shutdown. If that happens, we'll
            # just do an unclean shutdown.
            with contextlib.suppress(trio.BrokenResourceError, trio.BusyResourceError):
                await self._retry(self._ssl_object.unwrap, ignore_want_read=True)
        except:
            # Failure! Kill the stream and move on.
            await aclose_forcefully(self.transport_stream)
            raise
        else:
            # Success! Gracefully close the underlying stream.
            await self.transport_stream.aclose()
        finally:
            self._state = _State.CLOSED

    async def wait_send_all_might_not_block(self) -> None:
        """See :meth:`trio.abc.SendStream.wait_send_all_might_not_block`."""
        # This method's implementation is deceptively simple.
        #
        # First, we take the outer send lock, because of Trio's standard
        # semantics that wait_send_all_might_not_block and send_all
        # conflict.
        with self._outer_send_conflict_detector:
            self._check_status()
            # Then we take the inner send lock. We know that no other tasks
            # are calling self.send_all or self.wait_send_all_might_not_block,
            # because we have the outer_send_lock. But! There might be another
            # task calling self.receive_some -> transport_stream.send_all, in
            # which case if we were to call
            # transport_stream.wait_send_all_might_not_block directly we'd
            # have two tasks doing write-related operations on
            # transport_stream simultaneously, which is not allowed. We
            # *don't* want to raise this conflict to our caller, because it's
            # purely an internal affair – all they did was call
            # wait_send_all_might_not_block and receive_some at the same time,
            # which is totally valid. And waiting for the lock is OK, because
            # a call to send_all certainly wouldn't complete while the other
            # task holds the lock.
            async with self._inner_send_lock:
                # Now we have the lock, which creates another potential
                # problem: what if a call to self.receive_some attempts to do
                # transport_stream.send_all now? It'll have to wait for us to
                # finish! But that's OK, because we release the lock as soon
                # as the underlying stream becomes writable, and the
                # self.receive_some call wasn't going to make any progress
                # until then anyway.
                #
                # Of course, this does mean we might return *before* the
                # stream is logically writable, because immediately after we
                # return self.receive_some might write some data and make it
                # non-writable again. But that's OK too,
                # wait_send_all_might_not_block only guarantees that it
                # doesn't return late.
                await self.transport_stream.wait_send_all_might_not_block()


# this is necessary for Sphinx, see also `_abc.py`
SSLStream.__module__ = SSLStream.__module__.replace("._ssl", "")


@final
class SSLListener(Listener[SSLStream[T_Stream]]):
    """A :class:`~trio.abc.Listener` for SSL/TLS-encrypted servers.

    :class:`SSLListener` wraps around another Listener, and converts
    all incoming connections to encrypted connections by wrapping them
    in a :class:`SSLStream`.

    Args:
      transport_listener (~trio.abc.Listener): The listener whose incoming
          connections will be wrapped in :class:`SSLStream`.

      ssl_context (~ssl.SSLContext): The :class:`~ssl.SSLContext` that will be
          used for incoming connections.

      https_compatible (bool): Passed on to :class:`SSLStream`.

    Attributes:
      transport_listener (trio.abc.Listener): The underlying listener that was
          passed to ``__init__``.

    """

    def __init__(
        self,
        transport_listener: Listener[T_Stream],
        ssl_context: _stdlib_ssl.SSLContext,
        *,
        https_compatible: bool = False,
    ) -> None:
        self.transport_listener = transport_listener
        self._ssl_context = ssl_context
        self._https_compatible = https_compatible

    async def accept(self) -> SSLStream[T_Stream]:
        """Accept the next connection and wrap it in an :class:`SSLStream`.

        See :meth:`trio.abc.Listener.accept` for details.

        """
        transport_stream = await self.transport_listener.accept()
        return SSLStream(
            transport_stream,
            self._ssl_context,
            server_side=True,
            https_compatible=self._https_compatible,
        )

    async def aclose(self) -> None:
        """Close the transport listener."""
        await self.transport_listener.aclose()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\pygments\lexer.py ===
"""
    pygments.lexer
    ~~~~~~~~~~~~~~

    Base lexer classes.

    :copyright: Copyright 2006-2025 by the Pygments team, see AUTHORS.
    :license: BSD, see LICENSE for details.
"""

import re
import sys
import time

from pip._vendor.pygments.filter import apply_filters, Filter
from pip._vendor.pygments.filters import get_filter_by_name
from pip._vendor.pygments.token import Error, Text, Other, Whitespace, _TokenType
from pip._vendor.pygments.util import get_bool_opt, get_int_opt, get_list_opt, \
    make_analysator, Future, guess_decode
from pip._vendor.pygments.regexopt import regex_opt

__all__ = ['Lexer', 'RegexLexer', 'ExtendedRegexLexer', 'DelegatingLexer',
           'LexerContext', 'include', 'inherit', 'bygroups', 'using', 'this',
           'default', 'words', 'line_re']

line_re = re.compile('.*?\n')

_encoding_map = [(b'\xef\xbb\xbf', 'utf-8'),
                 (b'\xff\xfe\0\0', 'utf-32'),
                 (b'\0\0\xfe\xff', 'utf-32be'),
                 (b'\xff\xfe', 'utf-16'),
                 (b'\xfe\xff', 'utf-16be')]

_default_analyse = staticmethod(lambda x: 0.0)


class LexerMeta(type):
    """
    This metaclass automagically converts ``analyse_text`` methods into
    static methods which always return float values.
    """

    def __new__(mcs, name, bases, d):
        if 'analyse_text' in d:
            d['analyse_text'] = make_analysator(d['analyse_text'])
        return type.__new__(mcs, name, bases, d)


class Lexer(metaclass=LexerMeta):
    """
    Lexer for a specific language.

    See also :doc:`lexerdevelopment`, a high-level guide to writing
    lexers.

    Lexer classes have attributes used for choosing the most appropriate
    lexer based on various criteria.

    .. autoattribute:: name
       :no-value:
    .. autoattribute:: aliases
       :no-value:
    .. autoattribute:: filenames
       :no-value:
    .. autoattribute:: alias_filenames
    .. autoattribute:: mimetypes
       :no-value:
    .. autoattribute:: priority

    Lexers included in Pygments should have two additional attributes:

    .. autoattribute:: url
       :no-value:
    .. autoattribute:: version_added
       :no-value:

    Lexers included in Pygments may have additional attributes:

    .. autoattribute:: _example
       :no-value:

    You can pass options to the constructor. The basic options recognized
    by all lexers and processed by the base `Lexer` class are:

    ``stripnl``
        Strip leading and trailing newlines from the input (default: True).
    ``stripall``
        Strip all leading and trailing whitespace from the input
        (default: False).
    ``ensurenl``
        Make sure that the input ends with a newline (default: True).  This
        is required for some lexers that consume input linewise.

        .. versionadded:: 1.3

    ``tabsize``
        If given and greater than 0, expand tabs in the input (default: 0).
    ``encoding``
        If given, must be an encoding name. This encoding will be used to
        convert the input string to Unicode, if it is not already a Unicode
        string (default: ``'guess'``, which uses a simple UTF-8 / Locale /
        Latin1 detection.  Can also be ``'chardet'`` to use the chardet
        library, if it is installed.
    ``inencoding``
        Overrides the ``encoding`` if given.
    """

    #: Full name of the lexer, in human-readable form
    name = None

    #: A list of short, unique identifiers that can be used to look
    #: up the lexer from a list, e.g., using `get_lexer_by_name()`.
    aliases = []

    #: A list of `fnmatch` patterns that match filenames which contain
    #: content for this lexer. The patterns in this list should be unique among
    #: all lexers.
    filenames = []

    #: A list of `fnmatch` patterns that match filenames which may or may not
    #: contain content for this lexer. This list is used by the
    #: :func:`.guess_lexer_for_filename()` function, to determine which lexers
    #: are then included in guessing the correct one. That means that
    #: e.g. every lexer for HTML and a template language should include
    #: ``\*.html`` in this list.
    alias_filenames = []

    #: A list of MIME types for content that can be lexed with this lexer.
    mimetypes = []

    #: Priority, should multiple lexers match and no content is provided
    priority = 0

    #: URL of the language specification/definition. Used in the Pygments
    #: documentation. Set to an empty string to disable.
    url = None

    #: Version of Pygments in which the lexer was added.
    version_added = None

    #: Example file name. Relative to the ``tests/examplefiles`` directory.
    #: This is used by the documentation generator to show an example.
    _example = None

    def __init__(self, **options):
        """
        This constructor takes arbitrary options as keyword arguments.
        Every subclass must first process its own options and then call
        the `Lexer` constructor, since it processes the basic
        options like `stripnl`.

        An example looks like this:

        .. sourcecode:: python

           def __init__(self, **options):
               self.compress = options.get('compress', '')
               Lexer.__init__(self, **options)

        As these options must all be specifiable as strings (due to the
        command line usage), there are various utility functions
        available to help with that, see `Utilities`_.
        """
        self.options = options
        self.stripnl = get_bool_opt(options, 'stripnl', True)
        self.stripall = get_bool_opt(options, 'stripall', False)
        self.ensurenl = get_bool_opt(options, 'ensurenl', True)
        self.tabsize = get_int_opt(options, 'tabsize', 0)
        self.encoding = options.get('encoding', 'guess')
        self.encoding = options.get('inencoding') or self.encoding
        self.filters = []
        for filter_ in get_list_opt(options, 'filters', ()):
            self.add_filter(filter_)

    def __repr__(self):
        if self.options:
            return f'<pygments.lexers.{self.__class__.__name__} with {self.options!r}>'
        else:
            return f'<pygments.lexers.{self.__class__.__name__}>'

    def add_filter(self, filter_, **options):
        """
        Add a new stream filter to this lexer.
        """
        if not isinstance(filter_, Filter):
            filter_ = get_filter_by_name(filter_, **options)
        self.filters.append(filter_)

    def analyse_text(text):
        """
        A static method which is called for lexer guessing.

        It should analyse the text and return a float in the range
        from ``0.0`` to ``1.0``.  If it returns ``0.0``, the lexer
        will not be selected as the most probable one, if it returns
        ``1.0``, it will be selected immediately.  This is used by
        `guess_lexer`.

        The `LexerMeta` metaclass automatically wraps this function so
        that it works like a static method (no ``self`` or ``cls``
        parameter) and the return value is automatically converted to
        `float`. If the return value is an object that is boolean `False`
        it's the same as if the return values was ``0.0``.
        """

    def _preprocess_lexer_input(self, text):
        """Apply preprocessing such as decoding the input, removing BOM and normalizing newlines."""

        if not isinstance(text, str):
            if self.encoding == 'guess':
                text, _ = guess_decode(text)
            elif self.encoding == 'chardet':
                try:
                    # pip vendoring note: this code is not reachable by pip,
                    # removed import of chardet to make it clear.
                    raise ImportError('chardet is not vendored by pip')
                except ImportError as e:
                    raise ImportError('To enable chardet encoding guessing, '
                                      'please install the chardet library '
                                      'from http://chardet.feedparser.org/') from e
                # check for BOM first
                decoded = None
                for bom, encoding in _encoding_map:
                    if text.startswith(bom):
                        decoded = text[len(bom):].decode(encoding, 'replace')
                        break
                # no BOM found, so use chardet
                if decoded is None:
                    enc = chardet.detect(text[:1024])  # Guess using first 1KB
                    decoded = text.decode(enc.get('encoding') or 'utf-8',
                                          'replace')
                text = decoded
            else:
                text = text.decode(self.encoding)
                if text.startswith('\ufeff'):
                    text = text[len('\ufeff'):]
        else:
            if text.startswith('\ufeff'):
                text = text[len('\ufeff'):]

        # text now *is* a unicode string
        text = text.replace('\r\n', '\n')
        text = text.replace('\r', '\n')
        if self.stripall:
            text = text.strip()
        elif self.stripnl:
            text = text.strip('\n')
        if self.tabsize > 0:
            text = text.expandtabs(self.tabsize)
        if self.ensurenl and not text.endswith('\n'):
            text += '\n'

        return text

    def get_tokens(self, text, unfiltered=False):
        """
        This method is the basic interface of a lexer. It is called by
        the `highlight()` function. It must process the text and return an
        iterable of ``(tokentype, value)`` pairs from `text`.

        Normally, you don't need to override this method. The default
        implementation processes the options recognized by all lexers
        (`stripnl`, `stripall` and so on), and then yields all tokens
        from `get_tokens_unprocessed()`, with the ``index`` dropped.

        If `unfiltered` is set to `True`, the filtering mechanism is
        bypassed even if filters are defined.
        """
        text = self._preprocess_lexer_input(text)

        def streamer():
            for _, t, v in self.get_tokens_unprocessed(text):
                yield t, v
        stream = streamer()
        if not unfiltered:
            stream = apply_filters(stream, self.filters, self)
        return stream

    def get_tokens_unprocessed(self, text):
        """
        This method should process the text and return an iterable of
        ``(index, tokentype, value)`` tuples where ``index`` is the starting
        position of the token within the input text.

        It must be overridden by subclasses. It is recommended to
        implement it as a generator to maximize effectiveness.
        """
        raise NotImplementedError


class DelegatingLexer(Lexer):
    """
    This lexer takes two lexer as arguments. A root lexer and
    a language lexer. First everything is scanned using the language
    lexer, afterwards all ``Other`` tokens are lexed using the root
    lexer.

    The lexers from the ``template`` lexer package use this base lexer.
    """

    def __init__(self, _root_lexer, _language_lexer, _needle=Other, **options):
        self.root_lexer = _root_lexer(**options)
        self.language_lexer = _language_lexer(**options)
        self.needle = _needle
        Lexer.__init__(self, **options)

    def get_tokens_unprocessed(self, text):
        buffered = ''
        insertions = []
        lng_buffer = []
        for i, t, v in self.language_lexer.get_tokens_unprocessed(text):
            if t is self.needle:
                if lng_buffer:
                    insertions.append((len(buffered), lng_buffer))
                    lng_buffer = []
                buffered += v
            else:
                lng_buffer.append((i, t, v))
        if lng_buffer:
            insertions.append((len(buffered), lng_buffer))
        return do_insertions(insertions,
                             self.root_lexer.get_tokens_unprocessed(buffered))


# ------------------------------------------------------------------------------
# RegexLexer and ExtendedRegexLexer
#


class include(str):  # pylint: disable=invalid-name
    """
    Indicates that a state should include rules from another state.
    """
    pass


class _inherit:
    """
    Indicates the a state should inherit from its superclass.
    """
    def __repr__(self):
        return 'inherit'

inherit = _inherit()  # pylint: disable=invalid-name


class combined(tuple):  # pylint: disable=invalid-name
    """
    Indicates a state combined from multiple states.
    """

    def __new__(cls, *args):
        return tuple.__new__(cls, args)

    def __init__(self, *args):
        # tuple.__init__ doesn't do anything
        pass


class _PseudoMatch:
    """
    A pseudo match object constructed from a string.
    """

    def __init__(self, start, text):
        self._text = text
        self._start = start

    def start(self, arg=None):
        return self._start

    def end(self, arg=None):
        return self._start + len(self._text)

    def group(self, arg=None):
        if arg:
            raise IndexError('No such group')
        return self._text

    def groups(self):
        return (self._text,)

    def groupdict(self):
        return {}


def bygroups(*args):
    """
    Callback that yields multiple actions for each group in the match.
    """
    def callback(lexer, match, ctx=None):
        for i, action in enumerate(args):
            if action is None:
                continue
            elif type(action) is _TokenType:
                data = match.group(i + 1)
                if data:
                    yield match.start(i + 1), action, data
            else:
                data = match.group(i + 1)
                if data is not None:
                    if ctx:
                        ctx.pos = match.start(i + 1)
                    for item in action(lexer,
                                       _PseudoMatch(match.start(i + 1), data), ctx):
                        if item:
                            yield item
        if ctx:
            ctx.pos = match.end()
    return callback


class _This:
    """
    Special singleton used for indicating the caller class.
    Used by ``using``.
    """

this = _This()


def using(_other, **kwargs):
    """
    Callback that processes the match with a different lexer.

    The keyword arguments are forwarded to the lexer, except `state` which
    is handled separately.

    `state` specifies the state that the new lexer will start in, and can
    be an enumerable such as ('root', 'inline', 'string') or a simple
    string which is assumed to be on top of the root state.

    Note: For that to work, `_other` must not be an `ExtendedRegexLexer`.
    """
    gt_kwargs = {}
    if 'state' in kwargs:
        s = kwargs.pop('state')
        if isinstance(s, (list, tuple)):
            gt_kwargs['stack'] = s
        else:
            gt_kwargs['stack'] = ('root', s)

    if _other is this:
        def callback(lexer, match, ctx=None):
            # if keyword arguments are given the callback
            # function has to create a new lexer instance
            if kwargs:
                # XXX: cache that somehow
                kwargs.update(lexer.options)
                lx = lexer.__class__(**kwargs)
            else:
                lx = lexer
            s = match.start()
            for i, t, v in lx.get_tokens_unprocessed(match.group(), **gt_kwargs):
                yield i + s, t, v
            if ctx:
                ctx.pos = match.end()
    else:
        def callback(lexer, match, ctx=None):
            # XXX: cache that somehow
            kwargs.update(lexer.options)
            lx = _other(**kwargs)

            s = match.start()
            for i, t, v in lx.get_tokens_unprocessed(match.group(), **gt_kwargs):
                yield i + s, t, v
            if ctx:
                ctx.pos = match.end()
    return callback


class default:
    """
    Indicates a state or state action (e.g. #pop) to apply.
    For example default('#pop') is equivalent to ('', Token, '#pop')
    Note that state tuples may be used as well.

    .. versionadded:: 2.0
    """
    def __init__(self, state):
        self.state = state


class words(Future):
    """
    Indicates a list of literal words that is transformed into an optimized
    regex that matches any of the words.

    .. versionadded:: 2.0
    """
    def __init__(self, words, prefix='', suffix=''):
        self.words = words
        self.prefix = prefix
        self.suffix = suffix

    def get(self):
        return regex_opt(self.words, prefix=self.prefix, suffix=self.suffix)


class RegexLexerMeta(LexerMeta):
    """
    Metaclass for RegexLexer, creates the self._tokens attribute from
    self.tokens on the first instantiation.
    """

    def _process_regex(cls, regex, rflags, state):
        """Preprocess the regular expression component of a token definition."""
        if isinstance(regex, Future):
            regex = regex.get()
        return re.compile(regex, rflags).match

    def _process_token(cls, token):
        """Preprocess the token component of a token definition."""
        assert type(token) is _TokenType or callable(token), \
            f'token type must be simple type or callable, not {token!r}'
        return token

    def _process_new_state(cls, new_state, unprocessed, processed):
        """Preprocess the state transition action of a token definition."""
        if isinstance(new_state, str):
            # an existing state
            if new_state == '#pop':
                return -1
            elif new_state in unprocessed:
                return (new_state,)
            elif new_state == '#push':
                return new_state
            elif new_state[:5] == '#pop:':
                return -int(new_state[5:])
            else:
                assert False, f'unknown new state {new_state!r}'
        elif isinstance(new_state, combined):
            # combine a new state from existing ones
            tmp_state = '_tmp_%d' % cls._tmpname
            cls._tmpname += 1
            itokens = []
            for istate in new_state:
                assert istate != new_state, f'circular state ref {istate!r}'
                itokens.extend(cls._process_state(unprocessed,
                                                  processed, istate))
            processed[tmp_state] = itokens
            return (tmp_state,)
        elif isinstance(new_state, tuple):
            # push more than one state
            for istate in new_state:
                assert (istate in unprocessed or
                        istate in ('#pop', '#push')), \
                    'unknown new state ' + istate
            return new_state
        else:
            assert False, f'unknown new state def {new_state!r}'

    def _process_state(cls, unprocessed, processed, state):
        """Preprocess a single state definition."""
        assert isinstance(state, str), f"wrong state name {state!r}"
        assert state[0] != '#', f"invalid state name {state!r}"
        if state in processed:
            return processed[state]
        tokens = processed[state] = []
        rflags = cls.flags
        for tdef in unprocessed[state]:
            if isinstance(tdef, include):
                # it's a state reference
                assert tdef != state, f"circular state reference {state!r}"
                tokens.extend(cls._process_state(unprocessed, processed,
                                                 str(tdef)))
                continue
            if isinstance(tdef, _inherit):
                # should be processed already, but may not in the case of:
                # 1. the state has no counterpart in any parent
                # 2. the state includes more than one 'inherit'
                continue
            if isinstance(tdef, default):
                new_state = cls._process_new_state(tdef.state, unprocessed, processed)
                tokens.append((re.compile('').match, None, new_state))
                continue

            assert type(tdef) is tuple, f"wrong rule def {tdef!r}"

            try:
                rex = cls._process_regex(tdef[0], rflags, state)
            except Exception as err:
                raise ValueError(f"uncompilable regex {tdef[0]!r} in state {state!r} of {cls!r}: {err}") from err

            token = cls._process_token(tdef[1])

            if len(tdef) == 2:
                new_state = None
            else:
                new_state = cls._process_new_state(tdef[2],
                                                   unprocessed, processed)

            tokens.append((rex, token, new_state))
        return tokens

    def process_tokendef(cls, name, tokendefs=None):
        """Preprocess a dictionary of token definitions."""
        processed = cls._all_tokens[name] = {}
        tokendefs = tokendefs or cls.tokens[name]
        for state in list(tokendefs):
            cls._process_state(tokendefs, processed, state)
        return processed

    def get_tokendefs(cls):
        """
        Merge tokens from superclasses in MRO order, returning a single tokendef
        dictionary.

        Any state that is not defined by a subclass will be inherited
        automatically.  States that *are* defined by subclasses will, by
        default, override that state in the superclass.  If a subclass wishes to
        inherit definitions from a superclass, it can use the special value
        "inherit", which will cause the superclass' state definition to be
        included at that point in the state.
        """
        tokens = {}
        inheritable = {}
        for c in cls.__mro__:
            toks = c.__dict__.get('tokens', {})

            for state, items in toks.items():
                curitems = tokens.get(state)
                if curitems is None:
                    # N.b. because this is assigned by reference, sufficiently
                    # deep hierarchies are processed incrementally (e.g. for
                    # A(B), B(C), C(RegexLexer), B will be premodified so X(B)
                    # will not see any inherits in B).
                    tokens[state] = items
                    try:
                        inherit_ndx = items.index(inherit)
                    except ValueError:
                        continue
                    inheritable[state] = inherit_ndx
                    continue

                inherit_ndx = inheritable.pop(state, None)
                if inherit_ndx is None:
                    continue

                # Replace the "inherit" value with the items
                curitems[inherit_ndx:inherit_ndx+1] = items
                try:
                    # N.b. this is the index in items (that is, the superclass
                    # copy), so offset required when storing below.
                    new_inh_ndx = items.index(inherit)
                except ValueError:
                    pass
                else:
                    inheritable[state] = inherit_ndx + new_inh_ndx

        return tokens

    def __call__(cls, *args, **kwds):
        """Instantiate cls after preprocessing its token definitions."""
        if '_tokens' not in cls.__dict__:
            cls._all_tokens = {}
            cls._tmpname = 0
            if hasattr(cls, 'token_variants') and cls.token_variants:
                # don't process yet
                pass
            else:
                cls._tokens = cls.process_tokendef('', cls.get_tokendefs())

        return type.__call__(cls, *args, **kwds)


class RegexLexer(Lexer, metaclass=RegexLexerMeta):
    """
    Base for simple stateful regular expression-based lexers.
    Simplifies the lexing process so that you need only
    provide a list of states and regular expressions.
    """

    #: Flags for compiling the regular expressions.
    #: Defaults to MULTILINE.
    flags = re.MULTILINE

    #: At all time there is a stack of states. Initially, the stack contains
    #: a single state 'root'. The top of the stack is called "the current state".
    #:
    #: Dict of ``{'state': [(regex, tokentype, new_state), ...], ...}``
    #:
    #: ``new_state`` can be omitted to signify no state transition.
    #: If ``new_state`` is a string, it is pushed on the stack. This ensure
    #: the new current state is ``new_state``.
    #: If ``new_state`` is a tuple of strings, all of those strings are pushed
    #: on the stack and the current state will be the last element of the list.
    #: ``new_state`` can also be ``combined('state1', 'state2', ...)``
    #: to signify a new, anonymous state combined from the rules of two
    #: or more existing ones.
    #: Furthermore, it can be '#pop' to signify going back one step in
    #: the state stack, or '#push' to push the current state on the stack
    #: again. Note that if you push while in a combined state, the combined
    #: state itself is pushed, and not only the state in which the rule is
    #: defined.
    #:
    #: The tuple can also be replaced with ``include('state')``, in which
    #: case the rules from the state named by the string are included in the
    #: current one.
    tokens = {}

    def get_tokens_unprocessed(self, text, stack=('root',)):
        """
        Split ``text`` into (tokentype, text) pairs.

        ``stack`` is the initial stack (default: ``['root']``)
        """
        pos = 0
        tokendefs = self._tokens
        statestack = list(stack)
        statetokens = tokendefs[statestack[-1]]
        while 1:
            for rexmatch, action, new_state in statetokens:
                m = rexmatch(text, pos)
                if m:
                    if action is not None:
                        if type(action) is _TokenType:
                            yield pos, action, m.group()
                        else:
                            yield from action(self, m)
                    pos = m.end()
                    if new_state is not None:
                        # state transition
                        if isinstance(new_state, tuple):
                            for state in new_state:
                                if state == '#pop':
                                    if len(statestack) > 1:
                                        statestack.pop()
                                elif state == '#push':
                                    statestack.append(statestack[-1])
                                else:
                                    statestack.append(state)
                        elif isinstance(new_state, int):
                            # pop, but keep at least one state on the stack
                            # (random code leading to unexpected pops should
                            # not allow exceptions)
                            if abs(new_state) >= len(statestack):
                                del statestack[1:]
                            else:
                                del statestack[new_state:]
                        elif new_state == '#push':
                            statestack.append(statestack[-1])
                        else:
                            assert False, f"wrong state def: {new_state!r}"
                        statetokens = tokendefs[statestack[-1]]
                    break
            else:
                # We are here only if all state tokens have been considered
                # and there was not a match on any of them.
                try:
                    if text[pos] == '\n':
                        # at EOL, reset state to "root"
                        statestack = ['root']
                        statetokens = tokendefs['root']
                        yield pos, Whitespace, '\n'
                        pos += 1
                        continue
                    yield pos, Error, text[pos]
                    pos += 1
                except IndexError:
                    break


class LexerContext:
    """
    A helper object that holds lexer position data.
    """

    def __init__(self, text, pos, stack=None, end=None):
        self.text = text
        self.pos = pos
        self.end = end or len(text)  # end=0 not supported ;-)
        self.stack = stack or ['root']

    def __repr__(self):
        return f'LexerContext({self.text!r}, {self.pos!r}, {self.stack!r})'


class ExtendedRegexLexer(RegexLexer):
    """
    A RegexLexer that uses a context object to store its state.
    """

    def get_tokens_unprocessed(self, text=None, context=None):
        """
        Split ``text`` into (tokentype, text) pairs.
        If ``context`` is given, use this lexer context instead.
        """
        tokendefs = self._tokens
        if not context:
            ctx = LexerContext(text, 0)
            statetokens = tokendefs['root']
        else:
            ctx = context
            statetokens = tokendefs[ctx.stack[-1]]
            text = ctx.text
        while 1:
            for rexmatch, action, new_state in statetokens:
                m = rexmatch(text, ctx.pos, ctx.end)
                if m:
                    if action is not None:
                        if type(action) is _TokenType:
                            yield ctx.pos, action, m.group()
                            ctx.pos = m.end()
                        else:
                            yield from action(self, m, ctx)
                            if not new_state:
                                # altered the state stack?
                                statetokens = tokendefs[ctx.stack[-1]]
                    # CAUTION: callback must set ctx.pos!
                    if new_state is not None:
                        # state transition
                        if isinstance(new_state, tuple):
                            for state in new_state:
                                if state == '#pop':
                                    if len(ctx.stack) > 1:
                                        ctx.stack.pop()
                                elif state == '#push':
                                    ctx.stack.append(ctx.stack[-1])
                                else:
                                    ctx.stack.append(state)
                        elif isinstance(new_state, int):
                            # see RegexLexer for why this check is made
                            if abs(new_state) >= len(ctx.stack):
                                del ctx.stack[1:]
                            else:
                                del ctx.stack[new_state:]
                        elif new_state == '#push':
                            ctx.stack.append(ctx.stack[-1])
                        else:
                            assert False, f"wrong state def: {new_state!r}"
                        statetokens = tokendefs[ctx.stack[-1]]
                    break
            else:
                try:
                    if ctx.pos >= ctx.end:
                        break
                    if text[ctx.pos] == '\n':
                        # at EOL, reset state to "root"
                        ctx.stack = ['root']
                        statetokens = tokendefs['root']
                        yield ctx.pos, Text, '\n'
                        ctx.pos += 1
                        continue
                    yield ctx.pos, Error, text[ctx.pos]
                    ctx.pos += 1
                except IndexError:
                    break


def do_insertions(insertions, tokens):
    """
    Helper for lexers which must combine the results of several
    sublexers.

    ``insertions`` is a list of ``(index, itokens)`` pairs.
    Each ``itokens`` iterable should be inserted at position
    ``index`` into the token stream given by the ``tokens``
    argument.

    The result is a combined token stream.

    TODO: clean up the code here.
    """
    insertions = iter(insertions)
    try:
        index, itokens = next(insertions)
    except StopIteration:
        # no insertions
        yield from tokens
        return

    realpos = None
    insleft = True

    # iterate over the token stream where we want to insert
    # the tokens from the insertion list.
    for i, t, v in tokens:
        # first iteration. store the position of first item
        if realpos is None:
            realpos = i
        oldi = 0
        while insleft and i + len(v) >= index:
            tmpval = v[oldi:index - i]
            if tmpval:
                yield realpos, t, tmpval
                realpos += len(tmpval)
            for it_index, it_token, it_value in itokens:
                yield realpos, it_token, it_value
                realpos += len(it_value)
            oldi = index - i
            try:
                index, itokens = next(insertions)
            except StopIteration:
                insleft = False
                break  # not strictly necessary
        if oldi < len(v):
            yield realpos, t, v[oldi:]
            realpos += len(v) - oldi

    # leftover tokens
    while insleft:
        # no normal tokens, set realpos to zero
        realpos = realpos or 0
        for p, t, v in itokens:
            yield realpos, t, v
            realpos += len(v)
        try:
            index, itokens = next(insertions)
        except StopIteration:
            insleft = False
            break  # not strictly necessary


class ProfilingRegexLexerMeta(RegexLexerMeta):
    """Metaclass for ProfilingRegexLexer, collects regex timing info."""

    def _process_regex(cls, regex, rflags, state):
        if isinstance(regex, words):
            rex = regex_opt(regex.words, prefix=regex.prefix,
                            suffix=regex.suffix)
        else:
            rex = regex
        compiled = re.compile(rex, rflags)

        def match_func(text, pos, endpos=sys.maxsize):
            info = cls._prof_data[-1].setdefault((state, rex), [0, 0.0])
            t0 = time.time()
            res = compiled.match(text, pos, endpos)
            t1 = time.time()
            info[0] += 1
            info[1] += t1 - t0
            return res
        return match_func


class ProfilingRegexLexer(RegexLexer, metaclass=ProfilingRegexLexerMeta):
    """Drop-in replacement for RegexLexer that does profiling of its regexes."""

    _prof_data = []
    _prof_sort_index = 4  # defaults to time per call

    def get_tokens_unprocessed(self, text, stack=('root',)):
        # this needs to be a stack, since using(this) will produce nested calls
        self.__class__._prof_data.append({})
        yield from RegexLexer.get_tokens_unprocessed(self, text, stack)
        rawdata = self.__class__._prof_data.pop()
        data = sorted(((s, repr(r).strip('u\'').replace('\\\\', '\\')[:65],
                        n, 1000 * t, 1000 * t / n)
                       for ((s, r), (n, t)) in rawdata.items()),
                      key=lambda x: x[self._prof_sort_index],
                      reverse=True)
        sum_total = sum(x[3] for x in data)

        print()
        print('Profiling result for %s lexing %d chars in %.3f ms' %
              (self.__class__.__name__, len(text), sum_total))
        print('=' * 110)
        print('%-20s %-64s ncalls  tottime  percall' % ('state', 'regex'))
        print('-' * 110)
        for d in data:
            print('%-20s %-65s %5d %8.4f %8.4f' % d)
        print('=' * 110)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\io\formats\excel.py ===
"""
Utilities for conversion to writer-agnostic Excel representation.
"""
from __future__ import annotations

from collections.abc import (
    Hashable,
    Iterable,
    Mapping,
    Sequence,
)
import functools
import itertools
import re
from typing import (
    TYPE_CHECKING,
    Any,
    Callable,
    cast,
)
import warnings

import numpy as np

from pandas._libs.lib import is_list_like
from pandas.util._decorators import doc
from pandas.util._exceptions import find_stack_level

from pandas.core.dtypes import missing
from pandas.core.dtypes.common import (
    is_float,
    is_scalar,
)

from pandas import (
    DataFrame,
    Index,
    MultiIndex,
    PeriodIndex,
)
import pandas.core.common as com
from pandas.core.shared_docs import _shared_docs

from pandas.io.formats._color_data import CSS4_COLORS
from pandas.io.formats.css import (
    CSSResolver,
    CSSWarning,
)
from pandas.io.formats.format import get_level_lengths
from pandas.io.formats.printing import pprint_thing

if TYPE_CHECKING:
    from pandas._typing import (
        FilePath,
        IndexLabel,
        StorageOptions,
        WriteExcelBuffer,
    )

    from pandas import ExcelWriter


class ExcelCell:
    __fields__ = ("row", "col", "val", "style", "mergestart", "mergeend")
    __slots__ = __fields__

    def __init__(
        self,
        row: int,
        col: int,
        val,
        style=None,
        mergestart: int | None = None,
        mergeend: int | None = None,
    ) -> None:
        self.row = row
        self.col = col
        self.val = val
        self.style = style
        self.mergestart = mergestart
        self.mergeend = mergeend


class CssExcelCell(ExcelCell):
    def __init__(
        self,
        row: int,
        col: int,
        val,
        style: dict | None,
        css_styles: dict[tuple[int, int], list[tuple[str, Any]]] | None,
        css_row: int,
        css_col: int,
        css_converter: Callable | None,
        **kwargs,
    ) -> None:
        if css_styles and css_converter:
            # Use dict to get only one (case-insensitive) declaration per property
            declaration_dict = {
                prop.lower(): val for prop, val in css_styles[css_row, css_col]
            }
            # Convert to frozenset for order-invariant caching
            unique_declarations = frozenset(declaration_dict.items())
            style = css_converter(unique_declarations)

        super().__init__(row=row, col=col, val=val, style=style, **kwargs)


class CSSToExcelConverter:
    """
    A callable for converting CSS declarations to ExcelWriter styles

    Supports parts of CSS 2.2, with minimal CSS 3.0 support (e.g. text-shadow),
    focusing on font styling, backgrounds, borders and alignment.

    Operates by first computing CSS styles in a fairly generic
    way (see :meth:`compute_css`) then determining Excel style
    properties from CSS properties (see :meth:`build_xlstyle`).

    Parameters
    ----------
    inherited : str, optional
        CSS declarations understood to be the containing scope for the
        CSS processed by :meth:`__call__`.
    """

    NAMED_COLORS = CSS4_COLORS

    VERTICAL_MAP = {
        "top": "top",
        "text-top": "top",
        "middle": "center",
        "baseline": "bottom",
        "bottom": "bottom",
        "text-bottom": "bottom",
        # OpenXML also has 'justify', 'distributed'
    }

    BOLD_MAP = {
        "bold": True,
        "bolder": True,
        "600": True,
        "700": True,
        "800": True,
        "900": True,
        "normal": False,
        "lighter": False,
        "100": False,
        "200": False,
        "300": False,
        "400": False,
        "500": False,
    }

    ITALIC_MAP = {
        "normal": False,
        "italic": True,
        "oblique": True,
    }

    FAMILY_MAP = {
        "serif": 1,  # roman
        "sans-serif": 2,  # swiss
        "cursive": 4,  # script
        "fantasy": 5,  # decorative
    }

    BORDER_STYLE_MAP = {
        style.lower(): style
        for style in [
            "dashed",
            "mediumDashDot",
            "dashDotDot",
            "hair",
            "dotted",
            "mediumDashDotDot",
            "double",
            "dashDot",
            "slantDashDot",
            "mediumDashed",
        ]
    }

    # NB: Most of the methods here could be classmethods, as only __init__
    #     and __call__ make use of instance attributes.  We leave them as
    #     instancemethods so that users can easily experiment with extensions
    #     without monkey-patching.
    inherited: dict[str, str] | None

    def __init__(self, inherited: str | None = None) -> None:
        if inherited is not None:
            self.inherited = self.compute_css(inherited)
        else:
            self.inherited = None
        # We should avoid cache on the __call__ method.
        # Otherwise once the method __call__ has been called
        # garbage collection no longer deletes the instance.
        self._call_cached = functools.cache(self._call_uncached)

    compute_css = CSSResolver()

    def __call__(
        self, declarations: str | frozenset[tuple[str, str]]
    ) -> dict[str, dict[str, str]]:
        """
        Convert CSS declarations to ExcelWriter style.

        Parameters
        ----------
        declarations : str | frozenset[tuple[str, str]]
            CSS string or set of CSS declaration tuples.
            e.g. "font-weight: bold; background: blue" or
            {("font-weight", "bold"), ("background", "blue")}

        Returns
        -------
        xlstyle : dict
            A style as interpreted by ExcelWriter when found in
            ExcelCell.style.
        """
        return self._call_cached(declarations)

    def _call_uncached(
        self, declarations: str | frozenset[tuple[str, str]]
    ) -> dict[str, dict[str, str]]:
        properties = self.compute_css(declarations, self.inherited)
        return self.build_xlstyle(properties)

    def build_xlstyle(self, props: Mapping[str, str]) -> dict[str, dict[str, str]]:
        out = {
            "alignment": self.build_alignment(props),
            "border": self.build_border(props),
            "fill": self.build_fill(props),
            "font": self.build_font(props),
            "number_format": self.build_number_format(props),
        }

        # TODO: handle cell width and height: needs support in pandas.io.excel

        def remove_none(d: dict[str, str | None]) -> None:
            """Remove key where value is None, through nested dicts"""
            for k, v in list(d.items()):
                if v is None:
                    del d[k]
                elif isinstance(v, dict):
                    remove_none(v)
                    if not v:
                        del d[k]

        remove_none(out)
        return out

    def build_alignment(self, props: Mapping[str, str]) -> dict[str, bool | str | None]:
        # TODO: text-indent, padding-left -> alignment.indent
        return {
            "horizontal": props.get("text-align"),
            "vertical": self._get_vertical_alignment(props),
            "wrap_text": self._get_is_wrap_text(props),
        }

    def _get_vertical_alignment(self, props: Mapping[str, str]) -> str | None:
        vertical_align = props.get("vertical-align")
        if vertical_align:
            return self.VERTICAL_MAP.get(vertical_align)
        return None

    def _get_is_wrap_text(self, props: Mapping[str, str]) -> bool | None:
        if props.get("white-space") is None:
            return None
        return bool(props["white-space"] not in ("nowrap", "pre", "pre-line"))

    def build_border(
        self, props: Mapping[str, str]
    ) -> dict[str, dict[str, str | None]]:
        return {
            side: {
                "style": self._border_style(
                    props.get(f"border-{side}-style"),
                    props.get(f"border-{side}-width"),
                    self.color_to_excel(props.get(f"border-{side}-color")),
                ),
                "color": self.color_to_excel(props.get(f"border-{side}-color")),
            }
            for side in ["top", "right", "bottom", "left"]
        }

    def _border_style(self, style: str | None, width: str | None, color: str | None):
        # convert styles and widths to openxml, one of:
        #       'dashDot'
        #       'dashDotDot'
        #       'dashed'
        #       'dotted'
        #       'double'
        #       'hair'
        #       'medium'
        #       'mediumDashDot'
        #       'mediumDashDotDot'
        #       'mediumDashed'
        #       'slantDashDot'
        #       'thick'
        #       'thin'
        if width is None and style is None and color is None:
            # Return None will remove "border" from style dictionary
            return None

        if width is None and style is None:
            # Return "none" will keep "border" in style dictionary
            return "none"

        if style in ("none", "hidden"):
            return "none"

        width_name = self._get_width_name(width)
        if width_name is None:
            return "none"

        if style in (None, "groove", "ridge", "inset", "outset", "solid"):
            # not handled
            return width_name

        if style == "double":
            return "double"
        if style == "dotted":
            if width_name in ("hair", "thin"):
                return "dotted"
            return "mediumDashDotDot"
        if style == "dashed":
            if width_name in ("hair", "thin"):
                return "dashed"
            return "mediumDashed"
        elif style in self.BORDER_STYLE_MAP:
            # Excel-specific styles
            return self.BORDER_STYLE_MAP[style]
        else:
            warnings.warn(
                f"Unhandled border style format: {repr(style)}",
                CSSWarning,
                stacklevel=find_stack_level(),
            )
            return "none"

    def _get_width_name(self, width_input: str | None) -> str | None:
        width = self._width_to_float(width_input)
        if width < 1e-5:
            return None
        elif width < 1.3:
            return "thin"
        elif width < 2.8:
            return "medium"
        return "thick"

    def _width_to_float(self, width: str | None) -> float:
        if width is None:
            width = "2pt"
        return self._pt_to_float(width)

    def _pt_to_float(self, pt_string: str) -> float:
        assert pt_string.endswith("pt")
        return float(pt_string.rstrip("pt"))

    def build_fill(self, props: Mapping[str, str]):
        # TODO: perhaps allow for special properties
        #       -excel-pattern-bgcolor and -excel-pattern-type
        fill_color = props.get("background-color")
        if fill_color not in (None, "transparent", "none"):
            return {"fgColor": self.color_to_excel(fill_color), "patternType": "solid"}

    def build_number_format(self, props: Mapping[str, str]) -> dict[str, str | None]:
        fc = props.get("number-format")
        fc = fc.replace("§", ";") if isinstance(fc, str) else fc
        return {"format_code": fc}

    def build_font(
        self, props: Mapping[str, str]
    ) -> dict[str, bool | float | str | None]:
        font_names = self._get_font_names(props)
        decoration = self._get_decoration(props)
        return {
            "name": font_names[0] if font_names else None,
            "family": self._select_font_family(font_names),
            "size": self._get_font_size(props),
            "bold": self._get_is_bold(props),
            "italic": self._get_is_italic(props),
            "underline": ("single" if "underline" in decoration else None),
            "strike": ("line-through" in decoration) or None,
            "color": self.color_to_excel(props.get("color")),
            # shadow if nonzero digit before shadow color
            "shadow": self._get_shadow(props),
        }

    def _get_is_bold(self, props: Mapping[str, str]) -> bool | None:
        weight = props.get("font-weight")
        if weight:
            return self.BOLD_MAP.get(weight)
        return None

    def _get_is_italic(self, props: Mapping[str, str]) -> bool | None:
        font_style = props.get("font-style")
        if font_style:
            return self.ITALIC_MAP.get(font_style)
        return None

    def _get_decoration(self, props: Mapping[str, str]) -> Sequence[str]:
        decoration = props.get("text-decoration")
        if decoration is not None:
            return decoration.split()
        else:
            return ()

    def _get_underline(self, decoration: Sequence[str]) -> str | None:
        if "underline" in decoration:
            return "single"
        return None

    def _get_shadow(self, props: Mapping[str, str]) -> bool | None:
        if "text-shadow" in props:
            return bool(re.search("^[^#(]*[1-9]", props["text-shadow"]))
        return None

    def _get_font_names(self, props: Mapping[str, str]) -> Sequence[str]:
        font_names_tmp = re.findall(
            r"""(?x)
            (
            "(?:[^"]|\\")+"
            |
            '(?:[^']|\\')+'
            |
            [^'",]+
            )(?=,|\s*$)
        """,
            props.get("font-family", ""),
        )

        font_names = []
        for name in font_names_tmp:
            if name[:1] == '"':
                name = name[1:-1].replace('\\"', '"')
            elif name[:1] == "'":
                name = name[1:-1].replace("\\'", "'")
            else:
                name = name.strip()
            if name:
                font_names.append(name)
        return font_names

    def _get_font_size(self, props: Mapping[str, str]) -> float | None:
        size = props.get("font-size")
        if size is None:
            return size
        return self._pt_to_float(size)

    def _select_font_family(self, font_names: Sequence[str]) -> int | None:
        family = None
        for name in font_names:
            family = self.FAMILY_MAP.get(name)
            if family:
                break

        return family

    def color_to_excel(self, val: str | None) -> str | None:
        if val is None:
            return None

        if self._is_hex_color(val):
            return self._convert_hex_to_excel(val)

        try:
            return self.NAMED_COLORS[val]
        except KeyError:
            warnings.warn(
                f"Unhandled color format: {repr(val)}",
                CSSWarning,
                stacklevel=find_stack_level(),
            )
        return None

    def _is_hex_color(self, color_string: str) -> bool:
        return bool(color_string.startswith("#"))

    def _convert_hex_to_excel(self, color_string: str) -> str:
        code = color_string.lstrip("#")
        if self._is_shorthand_color(color_string):
            return (code[0] * 2 + code[1] * 2 + code[2] * 2).upper()
        else:
            return code.upper()

    def _is_shorthand_color(self, color_string: str) -> bool:
        """Check if color code is shorthand.

        #FFF is a shorthand as opposed to full #FFFFFF.
        """
        code = color_string.lstrip("#")
        if len(code) == 3:
            return True
        elif len(code) == 6:
            return False
        else:
            raise ValueError(f"Unexpected color {color_string}")


class ExcelFormatter:
    """
    Class for formatting a DataFrame to a list of ExcelCells,

    Parameters
    ----------
    df : DataFrame or Styler
    na_rep: na representation
    float_format : str, default None
        Format string for floating point numbers
    cols : sequence, optional
        Columns to write
    header : bool or sequence of str, default True
        Write out column names. If a list of string is given it is
        assumed to be aliases for the column names
    index : bool, default True
        output row names (index)
    index_label : str or sequence, default None
        Column label for index column(s) if desired. If None is given, and
        `header` and `index` are True, then the index names are used. A
        sequence should be given if the DataFrame uses MultiIndex.
    merge_cells : bool, default False
        Format MultiIndex and Hierarchical Rows as merged cells.
    inf_rep : str, default `'inf'`
        representation for np.inf values (which aren't representable in Excel)
        A `'-'` sign will be added in front of -inf.
    style_converter : callable, optional
        This translates Styler styles (CSS) into ExcelWriter styles.
        Defaults to ``CSSToExcelConverter()``.
        It should have signature css_declarations string -> excel style.
        This is only called for body cells.
    """

    max_rows = 2**20
    max_cols = 2**14

    def __init__(
        self,
        df,
        na_rep: str = "",
        float_format: str | None = None,
        cols: Sequence[Hashable] | None = None,
        header: Sequence[Hashable] | bool = True,
        index: bool = True,
        index_label: IndexLabel | None = None,
        merge_cells: bool = False,
        inf_rep: str = "inf",
        style_converter: Callable | None = None,
    ) -> None:
        self.rowcounter = 0
        self.na_rep = na_rep
        if not isinstance(df, DataFrame):
            self.styler = df
            self.styler._compute()  # calculate applied styles
            df = df.data
            if style_converter is None:
                style_converter = CSSToExcelConverter()
            self.style_converter: Callable | None = style_converter
        else:
            self.styler = None
            self.style_converter = None
        self.df = df
        if cols is not None:
            # all missing, raise
            if not len(Index(cols).intersection(df.columns)):
                raise KeyError("passes columns are not ALL present dataframe")

            if len(Index(cols).intersection(df.columns)) != len(set(cols)):
                # Deprecated in GH#17295, enforced in 1.0.0
                raise KeyError("Not all names specified in 'columns' are found")

            self.df = df.reindex(columns=cols)

        self.columns = self.df.columns
        self.float_format = float_format
        self.index = index
        self.index_label = index_label
        self.header = header
        self.merge_cells = merge_cells
        self.inf_rep = inf_rep

    @property
    def header_style(self) -> dict[str, dict[str, str | bool]]:
        return {
            "font": {"bold": True},
            "borders": {
                "top": "thin",
                "right": "thin",
                "bottom": "thin",
                "left": "thin",
            },
            "alignment": {"horizontal": "center", "vertical": "top"},
        }

    def _format_value(self, val):
        if is_scalar(val) and missing.isna(val):
            val = self.na_rep
        elif is_float(val):
            if missing.isposinf_scalar(val):
                val = self.inf_rep
            elif missing.isneginf_scalar(val):
                val = f"-{self.inf_rep}"
            elif self.float_format is not None:
                val = float(self.float_format % val)
        if getattr(val, "tzinfo", None) is not None:
            raise ValueError(
                "Excel does not support datetimes with "
                "timezones. Please ensure that datetimes "
                "are timezone unaware before writing to Excel."
            )
        return val

    def _format_header_mi(self) -> Iterable[ExcelCell]:
        if self.columns.nlevels > 1:
            if not self.index:
                raise NotImplementedError(
                    "Writing to Excel with MultiIndex columns and no "
                    "index ('index'=False) is not yet implemented."
                )

        if not (self._has_aliases or self.header):
            return

        columns = self.columns
        level_strs = columns._format_multi(
            sparsify=self.merge_cells, include_names=False
        )
        level_lengths = get_level_lengths(level_strs)
        coloffset = 0
        lnum = 0

        if self.index and isinstance(self.df.index, MultiIndex):
            coloffset = len(self.df.index[0]) - 1

        if self.merge_cells:
            # Format multi-index as a merged cells.
            for lnum, name in enumerate(columns.names):
                yield ExcelCell(
                    row=lnum,
                    col=coloffset,
                    val=name,
                    style=self.header_style,
                )

            for lnum, (spans, levels, level_codes) in enumerate(
                zip(level_lengths, columns.levels, columns.codes)
            ):
                values = levels.take(level_codes)
                for i, span_val in spans.items():
                    mergestart, mergeend = None, None
                    if span_val > 1:
                        mergestart, mergeend = lnum, coloffset + i + span_val
                    yield CssExcelCell(
                        row=lnum,
                        col=coloffset + i + 1,
                        val=values[i],
                        style=self.header_style,
                        css_styles=getattr(self.styler, "ctx_columns", None),
                        css_row=lnum,
                        css_col=i,
                        css_converter=self.style_converter,
                        mergestart=mergestart,
                        mergeend=mergeend,
                    )
        else:
            # Format in legacy format with dots to indicate levels.
            for i, values in enumerate(zip(*level_strs)):
                v = ".".join(map(pprint_thing, values))
                yield CssExcelCell(
                    row=lnum,
                    col=coloffset + i + 1,
                    val=v,
                    style=self.header_style,
                    css_styles=getattr(self.styler, "ctx_columns", None),
                    css_row=lnum,
                    css_col=i,
                    css_converter=self.style_converter,
                )

        self.rowcounter = lnum

    def _format_header_regular(self) -> Iterable[ExcelCell]:
        if self._has_aliases or self.header:
            coloffset = 0

            if self.index:
                coloffset = 1
                if isinstance(self.df.index, MultiIndex):
                    coloffset = len(self.df.index.names)

            colnames = self.columns
            if self._has_aliases:
                self.header = cast(Sequence, self.header)
                if len(self.header) != len(self.columns):
                    raise ValueError(
                        f"Writing {len(self.columns)} cols "
                        f"but got {len(self.header)} aliases"
                    )
                colnames = self.header

            for colindex, colname in enumerate(colnames):
                yield CssExcelCell(
                    row=self.rowcounter,
                    col=colindex + coloffset,
                    val=colname,
                    style=self.header_style,
                    css_styles=getattr(self.styler, "ctx_columns", None),
                    css_row=0,
                    css_col=colindex,
                    css_converter=self.style_converter,
                )

    def _format_header(self) -> Iterable[ExcelCell]:
        gen: Iterable[ExcelCell]

        if isinstance(self.columns, MultiIndex):
            gen = self._format_header_mi()
        else:
            gen = self._format_header_regular()

        gen2: Iterable[ExcelCell] = ()

        if self.df.index.names:
            row = [x if x is not None else "" for x in self.df.index.names] + [
                ""
            ] * len(self.columns)
            if functools.reduce(lambda x, y: x and y, (x != "" for x in row)):
                gen2 = (
                    ExcelCell(self.rowcounter, colindex, val, self.header_style)
                    for colindex, val in enumerate(row)
                )
                self.rowcounter += 1
        return itertools.chain(gen, gen2)

    def _format_body(self) -> Iterable[ExcelCell]:
        if isinstance(self.df.index, MultiIndex):
            return self._format_hierarchical_rows()
        else:
            return self._format_regular_rows()

    def _format_regular_rows(self) -> Iterable[ExcelCell]:
        if self._has_aliases or self.header:
            self.rowcounter += 1

        # output index and index_label?
        if self.index:
            # check aliases
            # if list only take first as this is not a MultiIndex
            if self.index_label and isinstance(
                self.index_label, (list, tuple, np.ndarray, Index)
            ):
                index_label = self.index_label[0]
            # if string good to go
            elif self.index_label and isinstance(self.index_label, str):
                index_label = self.index_label
            else:
                index_label = self.df.index.names[0]

            if isinstance(self.columns, MultiIndex):
                self.rowcounter += 1

            if index_label and self.header is not False:
                yield ExcelCell(self.rowcounter - 1, 0, index_label, self.header_style)

            # write index_values
            index_values = self.df.index
            if isinstance(self.df.index, PeriodIndex):
                index_values = self.df.index.to_timestamp()

            for idx, idxval in enumerate(index_values):
                yield CssExcelCell(
                    row=self.rowcounter + idx,
                    col=0,
                    val=idxval,
                    style=self.header_style,
                    css_styles=getattr(self.styler, "ctx_index", None),
                    css_row=idx,
                    css_col=0,
                    css_converter=self.style_converter,
                )
            coloffset = 1
        else:
            coloffset = 0

        yield from self._generate_body(coloffset)

    def _format_hierarchical_rows(self) -> Iterable[ExcelCell]:
        if self._has_aliases or self.header:
            self.rowcounter += 1

        gcolidx = 0

        if self.index:
            index_labels = self.df.index.names
            # check for aliases
            if self.index_label and isinstance(
                self.index_label, (list, tuple, np.ndarray, Index)
            ):
                index_labels = self.index_label

            # MultiIndex columns require an extra row
            # with index names (blank if None) for
            # unambiguous round-trip, unless not merging,
            # in which case the names all go on one row Issue #11328
            if isinstance(self.columns, MultiIndex) and self.merge_cells:
                self.rowcounter += 1

            # if index labels are not empty go ahead and dump
            if com.any_not_none(*index_labels) and self.header is not False:
                for cidx, name in enumerate(index_labels):
                    yield ExcelCell(self.rowcounter - 1, cidx, name, self.header_style)

            if self.merge_cells:
                # Format hierarchical rows as merged cells.
                level_strs = self.df.index._format_multi(
                    sparsify=True, include_names=False
                )
                level_lengths = get_level_lengths(level_strs)

                for spans, levels, level_codes in zip(
                    level_lengths, self.df.index.levels, self.df.index.codes
                ):
                    values = levels.take(
                        level_codes,
                        allow_fill=levels._can_hold_na,
                        fill_value=levels._na_value,
                    )

                    for i, span_val in spans.items():
                        mergestart, mergeend = None, None
                        if span_val > 1:
                            mergestart = self.rowcounter + i + span_val - 1
                            mergeend = gcolidx
                        yield CssExcelCell(
                            row=self.rowcounter + i,
                            col=gcolidx,
                            val=values[i],
                            style=self.header_style,
                            css_styles=getattr(self.styler, "ctx_index", None),
                            css_row=i,
                            css_col=gcolidx,
                            css_converter=self.style_converter,
                            mergestart=mergestart,
                            mergeend=mergeend,
                        )
                    gcolidx += 1

            else:
                # Format hierarchical rows with non-merged values.
                for indexcolvals in zip(*self.df.index):
                    for idx, indexcolval in enumerate(indexcolvals):
                        yield CssExcelCell(
                            row=self.rowcounter + idx,
                            col=gcolidx,
                            val=indexcolval,
                            style=self.header_style,
                            css_styles=getattr(self.styler, "ctx_index", None),
                            css_row=idx,
                            css_col=gcolidx,
                            css_converter=self.style_converter,
                        )
                    gcolidx += 1

        yield from self._generate_body(gcolidx)

    @property
    def _has_aliases(self) -> bool:
        """Whether the aliases for column names are present."""
        return is_list_like(self.header)

    def _generate_body(self, coloffset: int) -> Iterable[ExcelCell]:
        # Write the body of the frame data series by series.
        for colidx in range(len(self.columns)):
            series = self.df.iloc[:, colidx]
            for i, val in enumerate(series):
                yield CssExcelCell(
                    row=self.rowcounter + i,
                    col=colidx + coloffset,
                    val=val,
                    style=None,
                    css_styles=getattr(self.styler, "ctx", None),
                    css_row=i,
                    css_col=colidx,
                    css_converter=self.style_converter,
                )

    def get_formatted_cells(self) -> Iterable[ExcelCell]:
        for cell in itertools.chain(self._format_header(), self._format_body()):
            cell.val = self._format_value(cell.val)
            yield cell

    @doc(storage_options=_shared_docs["storage_options"])
    def write(
        self,
        writer: FilePath | WriteExcelBuffer | ExcelWriter,
        sheet_name: str = "Sheet1",
        startrow: int = 0,
        startcol: int = 0,
        freeze_panes: tuple[int, int] | None = None,
        engine: str | None = None,
        storage_options: StorageOptions | None = None,
        engine_kwargs: dict | None = None,
    ) -> None:
        """
        writer : path-like, file-like, or ExcelWriter object
            File path or existing ExcelWriter
        sheet_name : str, default 'Sheet1'
            Name of sheet which will contain DataFrame
        startrow :
            upper left cell row to dump data frame
        startcol :
            upper left cell column to dump data frame
        freeze_panes : tuple of integer (length 2), default None
            Specifies the one-based bottommost row and rightmost column that
            is to be frozen
        engine : string, default None
            write engine to use if writer is a path - you can also set this
            via the options ``io.excel.xlsx.writer``,
            or ``io.excel.xlsm.writer``.

        {storage_options}

        engine_kwargs: dict, optional
            Arbitrary keyword arguments passed to excel engine.
        """
        from pandas.io.excel import ExcelWriter

        num_rows, num_cols = self.df.shape
        if num_rows > self.max_rows or num_cols > self.max_cols:
            raise ValueError(
                f"This sheet is too large! Your sheet size is: {num_rows}, {num_cols} "
                f"Max sheet size is: {self.max_rows}, {self.max_cols}"
            )

        if engine_kwargs is None:
            engine_kwargs = {}

        formatted_cells = self.get_formatted_cells()
        if isinstance(writer, ExcelWriter):
            need_save = False
        else:
            writer = ExcelWriter(
                writer,
                engine=engine,
                storage_options=storage_options,
                engine_kwargs=engine_kwargs,
            )
            need_save = True

        try:
            writer._write_cells(
                formatted_cells,
                sheet_name,
                startrow=startrow,
                startcol=startcol,
                freeze_panes=freeze_panes,
            )
        finally:
            # make sure to close opened file handles
            if need_save:
                writer.close()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\ext\asyncio\result.py ===
# ext/asyncio/result.py
# Copyright (C) 2020-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
from __future__ import annotations

import operator
from typing import Any
from typing import AsyncIterator
from typing import Optional
from typing import overload
from typing import Sequence
from typing import Tuple
from typing import TYPE_CHECKING
from typing import TypeVar

from . import exc as async_exc
from ... import util
from ...engine import Result
from ...engine.result import _NO_ROW
from ...engine.result import _R
from ...engine.result import _WithKeys
from ...engine.result import FilterResult
from ...engine.result import FrozenResult
from ...engine.result import ResultMetaData
from ...engine.row import Row
from ...engine.row import RowMapping
from ...sql.base import _generative
from ...util.concurrency import greenlet_spawn
from ...util.typing import Literal
from ...util.typing import Self

if TYPE_CHECKING:
    from ...engine import CursorResult
    from ...engine.result import _KeyIndexType
    from ...engine.result import _UniqueFilterType

_T = TypeVar("_T", bound=Any)
_TP = TypeVar("_TP", bound=Tuple[Any, ...])


class AsyncCommon(FilterResult[_R]):
    __slots__ = ()

    _real_result: Result[Any]
    _metadata: ResultMetaData

    async def close(self) -> None:  # type: ignore[override]
        """Close this result."""

        await greenlet_spawn(self._real_result.close)

    @property
    def closed(self) -> bool:
        """proxies the .closed attribute of the underlying result object,
        if any, else raises ``AttributeError``.

        .. versionadded:: 2.0.0b3

        """
        return self._real_result.closed


class AsyncResult(_WithKeys, AsyncCommon[Row[_TP]]):
    """An asyncio wrapper around a :class:`_result.Result` object.

    The :class:`_asyncio.AsyncResult` only applies to statement executions that
    use a server-side cursor.  It is returned only from the
    :meth:`_asyncio.AsyncConnection.stream` and
    :meth:`_asyncio.AsyncSession.stream` methods.

    .. note:: As is the case with :class:`_engine.Result`, this object is
       used for ORM results returned by :meth:`_asyncio.AsyncSession.execute`,
       which can yield instances of ORM mapped objects either individually or
       within tuple-like rows.  Note that these result objects do not
       deduplicate instances or rows automatically as is the case with the
       legacy :class:`_orm.Query` object. For in-Python de-duplication of
       instances or rows, use the :meth:`_asyncio.AsyncResult.unique` modifier
       method.

    .. versionadded:: 1.4

    """

    __slots__ = ()

    _real_result: Result[_TP]

    def __init__(self, real_result: Result[_TP]):
        self._real_result = real_result

        self._metadata = real_result._metadata
        self._unique_filter_state = real_result._unique_filter_state
        self._source_supports_scalars = real_result._source_supports_scalars
        self._post_creational_filter = None

        # BaseCursorResult pre-generates the "_row_getter".  Use that
        # if available rather than building a second one
        if "_row_getter" in real_result.__dict__:
            self._set_memoized_attribute(
                "_row_getter", real_result.__dict__["_row_getter"]
            )

    @property
    def t(self) -> AsyncTupleResult[_TP]:
        """Apply a "typed tuple" typing filter to returned rows.

        The :attr:`_asyncio.AsyncResult.t` attribute is a synonym for
        calling the :meth:`_asyncio.AsyncResult.tuples` method.

        .. versionadded:: 2.0

        """
        return self  # type: ignore

    def tuples(self) -> AsyncTupleResult[_TP]:
        """Apply a "typed tuple" typing filter to returned rows.

        This method returns the same :class:`_asyncio.AsyncResult` object
        at runtime,
        however annotates as returning a :class:`_asyncio.AsyncTupleResult`
        object that will indicate to :pep:`484` typing tools that plain typed
        ``Tuple`` instances are returned rather than rows.  This allows
        tuple unpacking and ``__getitem__`` access of :class:`_engine.Row`
        objects to by typed, for those cases where the statement invoked
        itself included typing information.

        .. versionadded:: 2.0

        :return: the :class:`_result.AsyncTupleResult` type at typing time.

        .. seealso::

            :attr:`_asyncio.AsyncResult.t` - shorter synonym

            :attr:`_engine.Row.t` - :class:`_engine.Row` version

        """

        return self  # type: ignore

    @_generative
    def unique(self, strategy: Optional[_UniqueFilterType] = None) -> Self:
        """Apply unique filtering to the objects returned by this
        :class:`_asyncio.AsyncResult`.

        Refer to :meth:`_engine.Result.unique` in the synchronous
        SQLAlchemy API for a complete behavioral description.

        """
        self._unique_filter_state = (set(), strategy)
        return self

    def columns(self, *col_expressions: _KeyIndexType) -> Self:
        r"""Establish the columns that should be returned in each row.

        Refer to :meth:`_engine.Result.columns` in the synchronous
        SQLAlchemy API for a complete behavioral description.

        """
        return self._column_slices(col_expressions)

    async def partitions(
        self, size: Optional[int] = None
    ) -> AsyncIterator[Sequence[Row[_TP]]]:
        """Iterate through sub-lists of rows of the size given.

        An async iterator is returned::

            async def scroll_results(connection):
                result = await connection.stream(select(users_table))

                async for partition in result.partitions(100):
                    print("list of rows: %s" % partition)

        Refer to :meth:`_engine.Result.partitions` in the synchronous
        SQLAlchemy API for a complete behavioral description.

        """

        getter = self._manyrow_getter

        while True:
            partition = await greenlet_spawn(getter, self, size)
            if partition:
                yield partition
            else:
                break

    async def fetchall(self) -> Sequence[Row[_TP]]:
        """A synonym for the :meth:`_asyncio.AsyncResult.all` method.

        .. versionadded:: 2.0

        """

        return await greenlet_spawn(self._allrows)

    async def fetchone(self) -> Optional[Row[_TP]]:
        """Fetch one row.

        When all rows are exhausted, returns None.

        This method is provided for backwards compatibility with
        SQLAlchemy 1.x.x.

        To fetch the first row of a result only, use the
        :meth:`_asyncio.AsyncResult.first` method.  To iterate through all
        rows, iterate the :class:`_asyncio.AsyncResult` object directly.

        :return: a :class:`_engine.Row` object if no filters are applied,
         or ``None`` if no rows remain.

        """
        row = await greenlet_spawn(self._onerow_getter, self)
        if row is _NO_ROW:
            return None
        else:
            return row

    async def fetchmany(
        self, size: Optional[int] = None
    ) -> Sequence[Row[_TP]]:
        """Fetch many rows.

        When all rows are exhausted, returns an empty list.

        This method is provided for backwards compatibility with
        SQLAlchemy 1.x.x.

        To fetch rows in groups, use the
        :meth:`._asyncio.AsyncResult.partitions` method.

        :return: a list of :class:`_engine.Row` objects.

        .. seealso::

            :meth:`_asyncio.AsyncResult.partitions`

        """

        return await greenlet_spawn(self._manyrow_getter, self, size)

    async def all(self) -> Sequence[Row[_TP]]:
        """Return all rows in a list.

        Closes the result set after invocation.   Subsequent invocations
        will return an empty list.

        :return: a list of :class:`_engine.Row` objects.

        """

        return await greenlet_spawn(self._allrows)

    def __aiter__(self) -> AsyncResult[_TP]:
        return self

    async def __anext__(self) -> Row[_TP]:
        row = await greenlet_spawn(self._onerow_getter, self)
        if row is _NO_ROW:
            raise StopAsyncIteration()
        else:
            return row

    async def first(self) -> Optional[Row[_TP]]:
        """Fetch the first row or ``None`` if no row is present.

        Closes the result set and discards remaining rows.

        .. note::  This method returns one **row**, e.g. tuple, by default.
           To return exactly one single scalar value, that is, the first
           column of the first row, use the
           :meth:`_asyncio.AsyncResult.scalar` method,
           or combine :meth:`_asyncio.AsyncResult.scalars` and
           :meth:`_asyncio.AsyncResult.first`.

           Additionally, in contrast to the behavior of the legacy  ORM
           :meth:`_orm.Query.first` method, **no limit is applied** to the
           SQL query which was invoked to produce this
           :class:`_asyncio.AsyncResult`;
           for a DBAPI driver that buffers results in memory before yielding
           rows, all rows will be sent to the Python process and all but
           the first row will be discarded.

           .. seealso::

                :ref:`migration_20_unify_select`

        :return: a :class:`_engine.Row` object, or None
         if no rows remain.

        .. seealso::

            :meth:`_asyncio.AsyncResult.scalar`

            :meth:`_asyncio.AsyncResult.one`

        """
        return await greenlet_spawn(self._only_one_row, False, False, False)

    async def one_or_none(self) -> Optional[Row[_TP]]:
        """Return at most one result or raise an exception.

        Returns ``None`` if the result has no rows.
        Raises :class:`.MultipleResultsFound`
        if multiple rows are returned.

        .. versionadded:: 1.4

        :return: The first :class:`_engine.Row` or ``None`` if no row
         is available.

        :raises: :class:`.MultipleResultsFound`

        .. seealso::

            :meth:`_asyncio.AsyncResult.first`

            :meth:`_asyncio.AsyncResult.one`

        """
        return await greenlet_spawn(self._only_one_row, True, False, False)

    @overload
    async def scalar_one(self: AsyncResult[Tuple[_T]]) -> _T: ...

    @overload
    async def scalar_one(self) -> Any: ...

    async def scalar_one(self) -> Any:
        """Return exactly one scalar result or raise an exception.

        This is equivalent to calling :meth:`_asyncio.AsyncResult.scalars` and
        then :meth:`_asyncio.AsyncScalarResult.one`.

        .. seealso::

            :meth:`_asyncio.AsyncScalarResult.one`

            :meth:`_asyncio.AsyncResult.scalars`

        """
        return await greenlet_spawn(self._only_one_row, True, True, True)

    @overload
    async def scalar_one_or_none(
        self: AsyncResult[Tuple[_T]],
    ) -> Optional[_T]: ...

    @overload
    async def scalar_one_or_none(self) -> Optional[Any]: ...

    async def scalar_one_or_none(self) -> Optional[Any]:
        """Return exactly one scalar result or ``None``.

        This is equivalent to calling :meth:`_asyncio.AsyncResult.scalars` and
        then :meth:`_asyncio.AsyncScalarResult.one_or_none`.

        .. seealso::

            :meth:`_asyncio.AsyncScalarResult.one_or_none`

            :meth:`_asyncio.AsyncResult.scalars`

        """
        return await greenlet_spawn(self._only_one_row, True, False, True)

    async def one(self) -> Row[_TP]:
        """Return exactly one row or raise an exception.

        Raises :class:`.NoResultFound` if the result returns no
        rows, or :class:`.MultipleResultsFound` if multiple rows
        would be returned.

        .. note::  This method returns one **row**, e.g. tuple, by default.
           To return exactly one single scalar value, that is, the first
           column of the first row, use the
           :meth:`_asyncio.AsyncResult.scalar_one` method, or combine
           :meth:`_asyncio.AsyncResult.scalars` and
           :meth:`_asyncio.AsyncResult.one`.

        .. versionadded:: 1.4

        :return: The first :class:`_engine.Row`.

        :raises: :class:`.MultipleResultsFound`, :class:`.NoResultFound`

        .. seealso::

            :meth:`_asyncio.AsyncResult.first`

            :meth:`_asyncio.AsyncResult.one_or_none`

            :meth:`_asyncio.AsyncResult.scalar_one`

        """
        return await greenlet_spawn(self._only_one_row, True, True, False)

    @overload
    async def scalar(self: AsyncResult[Tuple[_T]]) -> Optional[_T]: ...

    @overload
    async def scalar(self) -> Any: ...

    async def scalar(self) -> Any:
        """Fetch the first column of the first row, and close the result set.

        Returns ``None`` if there are no rows to fetch.

        No validation is performed to test if additional rows remain.

        After calling this method, the object is fully closed,
        e.g. the :meth:`_engine.CursorResult.close`
        method will have been called.

        :return: a Python scalar value, or ``None`` if no rows remain.

        """
        return await greenlet_spawn(self._only_one_row, False, False, True)

    async def freeze(self) -> FrozenResult[_TP]:
        """Return a callable object that will produce copies of this
        :class:`_asyncio.AsyncResult` when invoked.

        The callable object returned is an instance of
        :class:`_engine.FrozenResult`.

        This is used for result set caching.  The method must be called
        on the result when it has been unconsumed, and calling the method
        will consume the result fully.   When the :class:`_engine.FrozenResult`
        is retrieved from a cache, it can be called any number of times where
        it will produce a new :class:`_engine.Result` object each time
        against its stored set of rows.

        .. seealso::

            :ref:`do_orm_execute_re_executing` - example usage within the
            ORM to implement a result-set cache.

        """

        return await greenlet_spawn(FrozenResult, self)

    @overload
    def scalars(
        self: AsyncResult[Tuple[_T]], index: Literal[0]
    ) -> AsyncScalarResult[_T]: ...

    @overload
    def scalars(self: AsyncResult[Tuple[_T]]) -> AsyncScalarResult[_T]: ...

    @overload
    def scalars(self, index: _KeyIndexType = 0) -> AsyncScalarResult[Any]: ...

    def scalars(self, index: _KeyIndexType = 0) -> AsyncScalarResult[Any]:
        """Return an :class:`_asyncio.AsyncScalarResult` filtering object which
        will return single elements rather than :class:`_row.Row` objects.

        Refer to :meth:`_result.Result.scalars` in the synchronous
        SQLAlchemy API for a complete behavioral description.

        :param index: integer or row key indicating the column to be fetched
         from each row, defaults to ``0`` indicating the first column.

        :return: a new :class:`_asyncio.AsyncScalarResult` filtering object
         referring to this :class:`_asyncio.AsyncResult` object.

        """
        return AsyncScalarResult(self._real_result, index)

    def mappings(self) -> AsyncMappingResult:
        """Apply a mappings filter to returned rows, returning an instance of
        :class:`_asyncio.AsyncMappingResult`.

        When this filter is applied, fetching rows will return
        :class:`_engine.RowMapping` objects instead of :class:`_engine.Row`
        objects.

        :return: a new :class:`_asyncio.AsyncMappingResult` filtering object
         referring to the underlying :class:`_result.Result` object.

        """

        return AsyncMappingResult(self._real_result)


class AsyncScalarResult(AsyncCommon[_R]):
    """A wrapper for a :class:`_asyncio.AsyncResult` that returns scalar values
    rather than :class:`_row.Row` values.

    The :class:`_asyncio.AsyncScalarResult` object is acquired by calling the
    :meth:`_asyncio.AsyncResult.scalars` method.

    Refer to the :class:`_result.ScalarResult` object in the synchronous
    SQLAlchemy API for a complete behavioral description.

    .. versionadded:: 1.4

    """

    __slots__ = ()

    _generate_rows = False

    def __init__(self, real_result: Result[Any], index: _KeyIndexType):
        self._real_result = real_result

        if real_result._source_supports_scalars:
            self._metadata = real_result._metadata
            self._post_creational_filter = None
        else:
            self._metadata = real_result._metadata._reduce([index])
            self._post_creational_filter = operator.itemgetter(0)

        self._unique_filter_state = real_result._unique_filter_state

    def unique(
        self,
        strategy: Optional[_UniqueFilterType] = None,
    ) -> Self:
        """Apply unique filtering to the objects returned by this
        :class:`_asyncio.AsyncScalarResult`.

        See :meth:`_asyncio.AsyncResult.unique` for usage details.

        """
        self._unique_filter_state = (set(), strategy)
        return self

    async def partitions(
        self, size: Optional[int] = None
    ) -> AsyncIterator[Sequence[_R]]:
        """Iterate through sub-lists of elements of the size given.

        Equivalent to :meth:`_asyncio.AsyncResult.partitions` except that
        scalar values, rather than :class:`_engine.Row` objects,
        are returned.

        """

        getter = self._manyrow_getter

        while True:
            partition = await greenlet_spawn(getter, self, size)
            if partition:
                yield partition
            else:
                break

    async def fetchall(self) -> Sequence[_R]:
        """A synonym for the :meth:`_asyncio.AsyncScalarResult.all` method."""

        return await greenlet_spawn(self._allrows)

    async def fetchmany(self, size: Optional[int] = None) -> Sequence[_R]:
        """Fetch many objects.

        Equivalent to :meth:`_asyncio.AsyncResult.fetchmany` except that
        scalar values, rather than :class:`_engine.Row` objects,
        are returned.

        """
        return await greenlet_spawn(self._manyrow_getter, self, size)

    async def all(self) -> Sequence[_R]:
        """Return all scalar values in a list.

        Equivalent to :meth:`_asyncio.AsyncResult.all` except that
        scalar values, rather than :class:`_engine.Row` objects,
        are returned.

        """
        return await greenlet_spawn(self._allrows)

    def __aiter__(self) -> AsyncScalarResult[_R]:
        return self

    async def __anext__(self) -> _R:
        row = await greenlet_spawn(self._onerow_getter, self)
        if row is _NO_ROW:
            raise StopAsyncIteration()
        else:
            return row

    async def first(self) -> Optional[_R]:
        """Fetch the first object or ``None`` if no object is present.

        Equivalent to :meth:`_asyncio.AsyncResult.first` except that
        scalar values, rather than :class:`_engine.Row` objects,
        are returned.

        """
        return await greenlet_spawn(self._only_one_row, False, False, False)

    async def one_or_none(self) -> Optional[_R]:
        """Return at most one object or raise an exception.

        Equivalent to :meth:`_asyncio.AsyncResult.one_or_none` except that
        scalar values, rather than :class:`_engine.Row` objects,
        are returned.

        """
        return await greenlet_spawn(self._only_one_row, True, False, False)

    async def one(self) -> _R:
        """Return exactly one object or raise an exception.

        Equivalent to :meth:`_asyncio.AsyncResult.one` except that
        scalar values, rather than :class:`_engine.Row` objects,
        are returned.

        """
        return await greenlet_spawn(self._only_one_row, True, True, False)


class AsyncMappingResult(_WithKeys, AsyncCommon[RowMapping]):
    """A wrapper for a :class:`_asyncio.AsyncResult` that returns dictionary
    values rather than :class:`_engine.Row` values.

    The :class:`_asyncio.AsyncMappingResult` object is acquired by calling the
    :meth:`_asyncio.AsyncResult.mappings` method.

    Refer to the :class:`_result.MappingResult` object in the synchronous
    SQLAlchemy API for a complete behavioral description.

    .. versionadded:: 1.4

    """

    __slots__ = ()

    _generate_rows = True

    _post_creational_filter = operator.attrgetter("_mapping")

    def __init__(self, result: Result[Any]):
        self._real_result = result
        self._unique_filter_state = result._unique_filter_state
        self._metadata = result._metadata
        if result._source_supports_scalars:
            self._metadata = self._metadata._reduce([0])

    def unique(
        self,
        strategy: Optional[_UniqueFilterType] = None,
    ) -> Self:
        """Apply unique filtering to the objects returned by this
        :class:`_asyncio.AsyncMappingResult`.

        See :meth:`_asyncio.AsyncResult.unique` for usage details.

        """
        self._unique_filter_state = (set(), strategy)
        return self

    def columns(self, *col_expressions: _KeyIndexType) -> Self:
        r"""Establish the columns that should be returned in each row."""
        return self._column_slices(col_expressions)

    async def partitions(
        self, size: Optional[int] = None
    ) -> AsyncIterator[Sequence[RowMapping]]:
        """Iterate through sub-lists of elements of the size given.

        Equivalent to :meth:`_asyncio.AsyncResult.partitions` except that
        :class:`_engine.RowMapping` values, rather than :class:`_engine.Row`
        objects, are returned.

        """

        getter = self._manyrow_getter

        while True:
            partition = await greenlet_spawn(getter, self, size)
            if partition:
                yield partition
            else:
                break

    async def fetchall(self) -> Sequence[RowMapping]:
        """A synonym for the :meth:`_asyncio.AsyncMappingResult.all` method."""

        return await greenlet_spawn(self._allrows)

    async def fetchone(self) -> Optional[RowMapping]:
        """Fetch one object.

        Equivalent to :meth:`_asyncio.AsyncResult.fetchone` except that
        :class:`_engine.RowMapping` values, rather than :class:`_engine.Row`
        objects, are returned.

        """

        row = await greenlet_spawn(self._onerow_getter, self)
        if row is _NO_ROW:
            return None
        else:
            return row

    async def fetchmany(
        self, size: Optional[int] = None
    ) -> Sequence[RowMapping]:
        """Fetch many rows.

        Equivalent to :meth:`_asyncio.AsyncResult.fetchmany` except that
        :class:`_engine.RowMapping` values, rather than :class:`_engine.Row`
        objects, are returned.

        """

        return await greenlet_spawn(self._manyrow_getter, self, size)

    async def all(self) -> Sequence[RowMapping]:
        """Return all rows in a list.

        Equivalent to :meth:`_asyncio.AsyncResult.all` except that
        :class:`_engine.RowMapping` values, rather than :class:`_engine.Row`
        objects, are returned.

        """

        return await greenlet_spawn(self._allrows)

    def __aiter__(self) -> AsyncMappingResult:
        return self

    async def __anext__(self) -> RowMapping:
        row = await greenlet_spawn(self._onerow_getter, self)
        if row is _NO_ROW:
            raise StopAsyncIteration()
        else:
            return row

    async def first(self) -> Optional[RowMapping]:
        """Fetch the first object or ``None`` if no object is present.

        Equivalent to :meth:`_asyncio.AsyncResult.first` except that
        :class:`_engine.RowMapping` values, rather than :class:`_engine.Row`
        objects, are returned.

        """
        return await greenlet_spawn(self._only_one_row, False, False, False)

    async def one_or_none(self) -> Optional[RowMapping]:
        """Return at most one object or raise an exception.

        Equivalent to :meth:`_asyncio.AsyncResult.one_or_none` except that
        :class:`_engine.RowMapping` values, rather than :class:`_engine.Row`
        objects, are returned.

        """
        return await greenlet_spawn(self._only_one_row, True, False, False)

    async def one(self) -> RowMapping:
        """Return exactly one object or raise an exception.

        Equivalent to :meth:`_asyncio.AsyncResult.one` except that
        :class:`_engine.RowMapping` values, rather than :class:`_engine.Row`
        objects, are returned.

        """
        return await greenlet_spawn(self._only_one_row, True, True, False)


class AsyncTupleResult(AsyncCommon[_R], util.TypingOnly):
    """A :class:`_asyncio.AsyncResult` that's typed as returning plain
    Python tuples instead of rows.

    Since :class:`_engine.Row` acts like a tuple in every way already,
    this class is a typing only class, regular :class:`_asyncio.AsyncResult` is
    still used at runtime.

    """

    __slots__ = ()

    if TYPE_CHECKING:

        async def partitions(
            self, size: Optional[int] = None
        ) -> AsyncIterator[Sequence[_R]]:
            """Iterate through sub-lists of elements of the size given.

            Equivalent to :meth:`_result.Result.partitions` except that
            tuple values, rather than :class:`_engine.Row` objects,
            are returned.

            """
            ...

        async def fetchone(self) -> Optional[_R]:
            """Fetch one tuple.

            Equivalent to :meth:`_result.Result.fetchone` except that
            tuple values, rather than :class:`_engine.Row`
            objects, are returned.

            """
            ...

        async def fetchall(self) -> Sequence[_R]:
            """A synonym for the :meth:`_engine.ScalarResult.all` method."""
            ...

        async def fetchmany(self, size: Optional[int] = None) -> Sequence[_R]:
            """Fetch many objects.

            Equivalent to :meth:`_result.Result.fetchmany` except that
            tuple values, rather than :class:`_engine.Row` objects,
            are returned.

            """
            ...

        async def all(self) -> Sequence[_R]:  # noqa: A001
            """Return all scalar values in a list.

            Equivalent to :meth:`_result.Result.all` except that
            tuple values, rather than :class:`_engine.Row` objects,
            are returned.

            """
            ...

        async def __aiter__(self) -> AsyncIterator[_R]: ...

        async def __anext__(self) -> _R: ...

        async def first(self) -> Optional[_R]:
            """Fetch the first object or ``None`` if no object is present.

            Equivalent to :meth:`_result.Result.first` except that
            tuple values, rather than :class:`_engine.Row` objects,
            are returned.


            """
            ...

        async def one_or_none(self) -> Optional[_R]:
            """Return at most one object or raise an exception.

            Equivalent to :meth:`_result.Result.one_or_none` except that
            tuple values, rather than :class:`_engine.Row` objects,
            are returned.

            """
            ...

        async def one(self) -> _R:
            """Return exactly one object or raise an exception.

            Equivalent to :meth:`_result.Result.one` except that
            tuple values, rather than :class:`_engine.Row` objects,
            are returned.

            """
            ...

        @overload
        async def scalar_one(self: AsyncTupleResult[Tuple[_T]]) -> _T: ...

        @overload
        async def scalar_one(self) -> Any: ...

        async def scalar_one(self) -> Any:
            """Return exactly one scalar result or raise an exception.

            This is equivalent to calling :meth:`_engine.Result.scalars`
            and then :meth:`_engine.AsyncScalarResult.one`.

            .. seealso::

                :meth:`_engine.AsyncScalarResult.one`

                :meth:`_engine.Result.scalars`

            """
            ...

        @overload
        async def scalar_one_or_none(
            self: AsyncTupleResult[Tuple[_T]],
        ) -> Optional[_T]: ...

        @overload
        async def scalar_one_or_none(self) -> Optional[Any]: ...

        async def scalar_one_or_none(self) -> Optional[Any]:
            """Return exactly one or no scalar result.

            This is equivalent to calling :meth:`_engine.Result.scalars`
            and then :meth:`_engine.AsyncScalarResult.one_or_none`.

            .. seealso::

                :meth:`_engine.AsyncScalarResult.one_or_none`

                :meth:`_engine.Result.scalars`

            """
            ...

        @overload
        async def scalar(
            self: AsyncTupleResult[Tuple[_T]],
        ) -> Optional[_T]: ...

        @overload
        async def scalar(self) -> Any: ...

        async def scalar(self) -> Any:
            """Fetch the first column of the first row, and close the result
            set.

            Returns ``None`` if there are no rows to fetch.

            No validation is performed to test if additional rows remain.

            After calling this method, the object is fully closed,
            e.g. the :meth:`_engine.CursorResult.close`
            method will have been called.

            :return: a Python scalar value , or ``None`` if no rows remain.

            """
            ...


_RT = TypeVar("_RT", bound="Result[Any]")


async def _ensure_sync_result(result: _RT, calling_method: Any) -> _RT:
    cursor_result: CursorResult[Any]

    try:
        is_cursor = result._is_cursor
    except AttributeError:
        # legacy execute(DefaultGenerator) case
        return result

    if not is_cursor:
        cursor_result = getattr(result, "raw", None)  # type: ignore
    else:
        cursor_result = result  # type: ignore
    if cursor_result and cursor_result.context._is_server_side:
        await greenlet_spawn(cursor_result.close)
        raise async_exc.AsyncMethodRequired(
            "Can't use the %s.%s() method with a "
            "server-side cursor. "
            "Use the %s.stream() method for an async "
            "streaming result set."
            % (
                calling_method.__self__.__class__.__name__,
                calling_method.__name__,
                calling_method.__self__.__class__.__name__,
            )
        )
    return result