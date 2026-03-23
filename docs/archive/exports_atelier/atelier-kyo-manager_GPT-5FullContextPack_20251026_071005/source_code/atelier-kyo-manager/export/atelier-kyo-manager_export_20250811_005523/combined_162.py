
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\matrices\domainscalar.py ===
"""

Module for the DomainScalar class.

A DomainScalar represents an element which is in a particular
Domain. The idea is that the DomainScalar class provides the
convenience routines for unifying elements with different domains.

It assists in Scalar Multiplication and getitem for DomainMatrix.

"""
from ..constructor import construct_domain

from sympy.polys.domains import Domain, ZZ


class DomainScalar:
    r"""
    docstring
    """

    def __new__(cls, element, domain):
        if not isinstance(domain, Domain):
            raise TypeError("domain should be of type Domain")
        if not domain.of_type(element):
            raise TypeError("element %s should be in domain %s" % (element, domain))
        return cls.new(element, domain)

    @classmethod
    def new(cls, element, domain):
        obj = super().__new__(cls)
        obj.element = element
        obj.domain = domain
        return obj

    def __repr__(self):
        return repr(self.element)

    @classmethod
    def from_sympy(cls, expr):
        [domain, [element]] = construct_domain([expr])
        return cls.new(element, domain)

    def to_sympy(self):
        return self.domain.to_sympy(self.element)

    def to_domain(self, domain):
        element = domain.convert_from(self.element, self.domain)
        return self.new(element, domain)

    def convert_to(self, domain):
        return self.to_domain(domain)

    def unify(self, other):
        domain = self.domain.unify(other.domain)
        return self.to_domain(domain), other.to_domain(domain)

    def __bool__(self):
        return bool(self.element)

    def __add__(self, other):
        if not isinstance(other, DomainScalar):
            return NotImplemented
        self, other = self.unify(other)
        return self.new(self.element + other.element, self.domain)

    def __sub__(self, other):
        if not isinstance(other, DomainScalar):
            return NotImplemented
        self, other = self.unify(other)
        return self.new(self.element - other.element, self.domain)

    def __mul__(self, other):
        if not isinstance(other, DomainScalar):
            if isinstance(other, int):
                other = DomainScalar(ZZ(other), ZZ)
            else:
                return NotImplemented

        self, other = self.unify(other)
        return self.new(self.element * other.element, self.domain)

    def __floordiv__(self, other):
        if not isinstance(other, DomainScalar):
            return NotImplemented
        self, other = self.unify(other)
        return self.new(self.domain.quo(self.element, other.element), self.domain)

    def __mod__(self, other):
        if not isinstance(other, DomainScalar):
            return NotImplemented
        self, other = self.unify(other)
        return self.new(self.domain.rem(self.element, other.element), self.domain)

    def __divmod__(self, other):
        if not isinstance(other, DomainScalar):
            return NotImplemented
        self, other = self.unify(other)
        q, r = self.domain.div(self.element, other.element)
        return (self.new(q, self.domain), self.new(r, self.domain))

    def __pow__(self, n):
        if not isinstance(n, int):
            return NotImplemented
        return self.new(self.element**n, self.domain)

    def __pos__(self):
        return self.new(+self.element, self.domain)

    def __neg__(self):
        return self.new(-self.element, self.domain)

    def __eq__(self, other):
        if not isinstance(other, DomainScalar):
            return NotImplemented
        return self.element == other.element and self.domain == other.domain

    def is_zero(self):
        return self.element == self.domain.zero

    def is_one(self):
        return self.element == self.domain.one

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\utilities\mathml\__init__.py ===
"""Module with some functions for MathML, like transforming MathML
content in MathML presentation.

To use this module, you will need lxml.
"""

from pathlib import Path

from sympy.utilities.decorator import doctest_depends_on


__doctest_requires__ = {('apply_xsl', 'c2p'): ['lxml']}


def add_mathml_headers(s):
    return """<math xmlns:mml="http://www.w3.org/1998/Math/MathML"
      xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
      xsi:schemaLocation="http://www.w3.org/1998/Math/MathML
        http://www.w3.org/Math/XMLSchema/mathml2/mathml2.xsd">""" + s + "</math>"


def _read_binary(pkgname, filename):
    import sys

    if sys.version_info >= (3, 10):
        # files was added in Python 3.9 but only seems to work here in 3.10+
        from importlib.resources import files
        return files(pkgname).joinpath(filename).read_bytes()
    else:
        # read_binary was deprecated in Python 3.11
        from importlib.resources import read_binary
        return read_binary(pkgname, filename)


def _read_xsl(xsl):
    # Previously these values were allowed:
    if xsl == 'mathml/data/simple_mmlctop.xsl':
        xsl = 'simple_mmlctop.xsl'
    elif xsl == 'mathml/data/mmlctop.xsl':
        xsl = 'mmlctop.xsl'
    elif xsl == 'mathml/data/mmltex.xsl':
        xsl = 'mmltex.xsl'

    if xsl in ['simple_mmlctop.xsl', 'mmlctop.xsl', 'mmltex.xsl']:
        xslbytes = _read_binary('sympy.utilities.mathml.data', xsl)
    else:
        xslbytes = Path(xsl).read_bytes()

    return xslbytes


@doctest_depends_on(modules=('lxml',))
def apply_xsl(mml, xsl):
    """Apply a xsl to a MathML string.

    Parameters
    ==========

    mml
        A string with MathML code.
    xsl
        A string giving the name of an xsl (xml stylesheet) file which can be
        found in sympy/utilities/mathml/data. The following files are supplied
        with SymPy:

        - mmlctop.xsl
        - mmltex.xsl
        - simple_mmlctop.xsl

        Alternatively, a full path to an xsl file can be given.

    Examples
    ========

    >>> from sympy.utilities.mathml import apply_xsl
    >>> xsl = 'simple_mmlctop.xsl'
    >>> mml = '<apply> <plus/> <ci>a</ci> <ci>b</ci> </apply>'
    >>> res = apply_xsl(mml,xsl)
    >>> print(res)
    <?xml version="1.0"?>
    <mrow xmlns="http://www.w3.org/1998/Math/MathML">
      <mi>a</mi>
      <mo> + </mo>
      <mi>b</mi>
    </mrow>
    """
    from lxml import etree

    parser = etree.XMLParser(resolve_entities=False)
    ac = etree.XSLTAccessControl.DENY_ALL

    s = etree.XML(_read_xsl(xsl), parser=parser)
    transform = etree.XSLT(s, access_control=ac)
    doc = etree.XML(mml, parser=parser)
    result = transform(doc)
    s = str(result)
    return s


@doctest_depends_on(modules=('lxml',))
def c2p(mml, simple=False):
    """Transforms a document in MathML content (like the one that sympy produces)
    in one document in MathML presentation, more suitable for printing, and more
    widely accepted

    Examples
    ========

    >>> from sympy.utilities.mathml import c2p
    >>> mml = '<apply> <exp/> <cn>2</cn> </apply>'
    >>> c2p(mml,simple=True) != c2p(mml,simple=False)
    True

    """

    if not mml.startswith('<math'):
        mml = add_mathml_headers(mml)

    if simple:
        return apply_xsl(mml, 'mathml/data/simple_mmlctop.xsl')

    return apply_xsl(mml, 'mathml/data/mmlctop.xsl')

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\tools\test_to_datetime.py ===
""" test to_datetime """

import calendar
from collections import deque
from datetime import (
    date,
    datetime,
    timedelta,
    timezone,
)
from decimal import Decimal
import locale

from dateutil.parser import parse
from dateutil.tz.tz import tzoffset
import numpy as np
import pytest
import pytz

from pandas._libs import tslib
from pandas._libs.tslibs import (
    iNaT,
    parsing,
)
from pandas.errors import (
    OutOfBoundsDatetime,
    OutOfBoundsTimedelta,
)
import pandas.util._test_decorators as td

from pandas.core.dtypes.common import is_datetime64_ns_dtype

import pandas as pd
from pandas import (
    DataFrame,
    DatetimeIndex,
    Index,
    NaT,
    Series,
    Timestamp,
    date_range,
    isna,
    to_datetime,
)
import pandas._testing as tm
from pandas.core.arrays import DatetimeArray
from pandas.core.tools import datetimes as tools
from pandas.core.tools.datetimes import start_caching_at

PARSING_ERR_MSG = (
    r"You might want to try:\n"
    r"    - passing `format` if your strings have a consistent format;\n"
    r"    - passing `format=\'ISO8601\'` if your strings are all ISO8601 "
    r"but not necessarily in exactly the same format;\n"
    r"    - passing `format=\'mixed\'`, and the format will be inferred "
    r"for each element individually. You might want to use `dayfirst` "
    r"alongside this."
)

pytestmark = pytest.mark.filterwarnings(
    "ignore:errors='ignore' is deprecated:FutureWarning"
)


@pytest.fixture(params=[True, False])
def cache(request):
    """
    cache keyword to pass to to_datetime.
    """
    return request.param


class TestTimeConversionFormats:
    @pytest.mark.parametrize("readonly", [True, False])
    def test_to_datetime_readonly(self, readonly):
        # GH#34857
        arr = np.array([], dtype=object)
        if readonly:
            arr.setflags(write=False)
        result = to_datetime(arr)
        expected = to_datetime([])
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "format, expected",
        [
            [
                "%d/%m/%Y",
                [Timestamp("20000101"), Timestamp("20000201"), Timestamp("20000301")],
            ],
            [
                "%m/%d/%Y",
                [Timestamp("20000101"), Timestamp("20000102"), Timestamp("20000103")],
            ],
        ],
    )
    def test_to_datetime_format(self, cache, index_or_series, format, expected):
        values = index_or_series(["1/1/2000", "1/2/2000", "1/3/2000"])
        result = to_datetime(values, format=format, cache=cache)
        expected = index_or_series(expected)
        tm.assert_equal(result, expected)

    @pytest.mark.parametrize(
        "arg, expected, format",
        [
            ["1/1/2000", "20000101", "%d/%m/%Y"],
            ["1/1/2000", "20000101", "%m/%d/%Y"],
            ["1/2/2000", "20000201", "%d/%m/%Y"],
            ["1/2/2000", "20000102", "%m/%d/%Y"],
            ["1/3/2000", "20000301", "%d/%m/%Y"],
            ["1/3/2000", "20000103", "%m/%d/%Y"],
        ],
    )
    def test_to_datetime_format_scalar(self, cache, arg, expected, format):
        result = to_datetime(arg, format=format, cache=cache)
        expected = Timestamp(expected)
        assert result == expected

    def test_to_datetime_format_YYYYMMDD(self, cache):
        ser = Series([19801222, 19801222] + [19810105] * 5)
        expected = Series([Timestamp(x) for x in ser.apply(str)])

        result = to_datetime(ser, format="%Y%m%d", cache=cache)
        tm.assert_series_equal(result, expected)

        result = to_datetime(ser.apply(str), format="%Y%m%d", cache=cache)
        tm.assert_series_equal(result, expected)

    def test_to_datetime_format_YYYYMMDD_with_nat(self, cache):
        # Explicit cast to float to explicit cast when setting np.nan
        ser = Series([19801222, 19801222] + [19810105] * 5, dtype="float")
        # with NaT
        expected = Series(
            [Timestamp("19801222"), Timestamp("19801222")] + [Timestamp("19810105")] * 5
        )
        expected[2] = np.nan
        ser[2] = np.nan

        result = to_datetime(ser, format="%Y%m%d", cache=cache)
        tm.assert_series_equal(result, expected)

        # string with NaT
        ser2 = ser.apply(str)
        ser2[2] = "nat"
        with pytest.raises(
            ValueError,
            match=(
                'unconverted data remains when parsing with format "%Y%m%d": ".0", '
                "at position 0"
            ),
        ):
            # https://github.com/pandas-dev/pandas/issues/50051
            to_datetime(ser2, format="%Y%m%d", cache=cache)

    def test_to_datetime_format_YYYYMM_with_nat(self, cache):
        # https://github.com/pandas-dev/pandas/issues/50237
        # Explicit cast to float to explicit cast when setting np.nan
        ser = Series([198012, 198012] + [198101] * 5, dtype="float")
        expected = Series(
            [Timestamp("19801201"), Timestamp("19801201")] + [Timestamp("19810101")] * 5
        )
        expected[2] = np.nan
        ser[2] = np.nan
        result = to_datetime(ser, format="%Y%m", cache=cache)
        tm.assert_series_equal(result, expected)

    def test_to_datetime_format_YYYYMMDD_ignore(self, cache):
        # coercion
        # GH 7930, GH 14487
        ser = Series([20121231, 20141231, 99991231])
        result = to_datetime(ser, format="%Y%m%d", errors="ignore", cache=cache)
        expected = Series(
            [20121231, 20141231, 99991231],
            dtype=object,
        )
        tm.assert_series_equal(result, expected)

    def test_to_datetime_format_YYYYMMDD_ignore_with_outofbounds(self, cache):
        # https://github.com/pandas-dev/pandas/issues/26493
        result = to_datetime(
            ["15010101", "20150101", np.nan],
            format="%Y%m%d",
            errors="ignore",
            cache=cache,
        )
        expected = Index(["15010101", "20150101", np.nan], dtype=object)
        tm.assert_index_equal(result, expected)

    def test_to_datetime_format_YYYYMMDD_coercion(self, cache):
        # coercion
        # GH 7930
        ser = Series([20121231, 20141231, 99991231])
        result = to_datetime(ser, format="%Y%m%d", errors="coerce", cache=cache)
        expected = Series(["20121231", "20141231", "NaT"], dtype="M8[ns]")
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize(
        "input_s",
        [
            # Null values with Strings
            ["19801222", "20010112", None],
            ["19801222", "20010112", np.nan],
            ["19801222", "20010112", NaT],
            ["19801222", "20010112", "NaT"],
            # Null values with Integers
            [19801222, 20010112, None],
            [19801222, 20010112, np.nan],
            [19801222, 20010112, NaT],
            [19801222, 20010112, "NaT"],
        ],
    )
    def test_to_datetime_format_YYYYMMDD_with_none(self, input_s):
        # GH 30011
        # format='%Y%m%d'
        # with None
        expected = Series([Timestamp("19801222"), Timestamp("20010112"), NaT])
        result = Series(to_datetime(input_s, format="%Y%m%d"))
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize(
        "input_s, expected",
        [
            # NaN before strings with invalid date values
            [
                Series(["19801222", np.nan, "20010012", "10019999"]),
                Series([Timestamp("19801222"), np.nan, np.nan, np.nan]),
            ],
            # NaN after strings with invalid date values
            [
                Series(["19801222", "20010012", "10019999", np.nan]),
                Series([Timestamp("19801222"), np.nan, np.nan, np.nan]),
            ],
            # NaN before integers with invalid date values
            [
                Series([20190813, np.nan, 20010012, 20019999]),
                Series([Timestamp("20190813"), np.nan, np.nan, np.nan]),
            ],
            # NaN after integers with invalid date values
            [
                Series([20190813, 20010012, np.nan, 20019999]),
                Series([Timestamp("20190813"), np.nan, np.nan, np.nan]),
            ],
        ],
    )
    def test_to_datetime_format_YYYYMMDD_overflow(self, input_s, expected):
        # GH 25512
        # format='%Y%m%d', errors='coerce'
        result = to_datetime(input_s, format="%Y%m%d", errors="coerce")
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize(
        "data, format, expected",
        [
            ([pd.NA], "%Y%m%d%H%M%S", DatetimeIndex(["NaT"])),
            ([pd.NA], None, DatetimeIndex(["NaT"])),
            (
                [pd.NA, "20210202202020"],
                "%Y%m%d%H%M%S",
                DatetimeIndex(["NaT", "2021-02-02 20:20:20"]),
            ),
            (["201010", pd.NA], "%y%m%d", DatetimeIndex(["2020-10-10", "NaT"])),
            (["201010", pd.NA], "%d%m%y", DatetimeIndex(["2010-10-20", "NaT"])),
            ([None, np.nan, pd.NA], None, DatetimeIndex(["NaT", "NaT", "NaT"])),
            ([None, np.nan, pd.NA], "%Y%m%d", DatetimeIndex(["NaT", "NaT", "NaT"])),
        ],
    )
    def test_to_datetime_with_NA(self, data, format, expected):
        # GH#42957
        result = to_datetime(data, format=format)
        tm.assert_index_equal(result, expected)

    def test_to_datetime_with_NA_with_warning(self):
        # GH#42957
        result = to_datetime(["201010", pd.NA])
        expected = DatetimeIndex(["2010-10-20", "NaT"])
        tm.assert_index_equal(result, expected)

    def test_to_datetime_format_integer(self, cache):
        # GH 10178
        ser = Series([2000, 2001, 2002])
        expected = Series([Timestamp(x) for x in ser.apply(str)])

        result = to_datetime(ser, format="%Y", cache=cache)
        tm.assert_series_equal(result, expected)

        ser = Series([200001, 200105, 200206])
        expected = Series([Timestamp(x[:4] + "-" + x[4:]) for x in ser.apply(str)])

        result = to_datetime(ser, format="%Y%m", cache=cache)
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize(
        "int_date, expected",
        [
            # valid date, length == 8
            [20121030, datetime(2012, 10, 30)],
            # short valid date, length == 6
            [199934, datetime(1999, 3, 4)],
            # long integer date partially parsed to datetime(2012,1,1), length > 8
            [2012010101, 2012010101],
            # invalid date partially parsed to datetime(2012,9,9), length == 8
            [20129930, 20129930],
            # short integer date partially parsed to datetime(2012,9,9), length < 8
            [2012993, 2012993],
            # short invalid date, length == 4
            [2121, 2121],
        ],
    )
    def test_int_to_datetime_format_YYYYMMDD_typeerror(self, int_date, expected):
        # GH 26583
        result = to_datetime(int_date, format="%Y%m%d", errors="ignore")
        assert result == expected

    def test_to_datetime_format_microsecond(self, cache):
        month_abbr = calendar.month_abbr[4]
        val = f"01-{month_abbr}-2011 00:00:01.978"

        format = "%d-%b-%Y %H:%M:%S.%f"
        result = to_datetime(val, format=format, cache=cache)
        exp = datetime.strptime(val, format)
        assert result == exp

    @pytest.mark.parametrize(
        "value, format, dt",
        [
            ["01/10/2010 15:20", "%m/%d/%Y %H:%M", Timestamp("2010-01-10 15:20")],
            ["01/10/2010 05:43", "%m/%d/%Y %I:%M", Timestamp("2010-01-10 05:43")],
            [
                "01/10/2010 13:56:01",
                "%m/%d/%Y %H:%M:%S",
                Timestamp("2010-01-10 13:56:01"),
            ],
            # The 3 tests below are locale-dependent.
            # They pass, except when the machine locale is zh_CN or it_IT .
            pytest.param(
                "01/10/2010 08:14 PM",
                "%m/%d/%Y %I:%M %p",
                Timestamp("2010-01-10 20:14"),
                marks=pytest.mark.xfail(
                    locale.getlocale()[0] in ("zh_CN", "it_IT"),
                    reason="fail on a CI build with LC_ALL=zh_CN.utf8/it_IT.utf8",
                    strict=False,
                ),
            ),
            pytest.param(
                "01/10/2010 07:40 AM",
                "%m/%d/%Y %I:%M %p",
                Timestamp("2010-01-10 07:40"),
                marks=pytest.mark.xfail(
                    locale.getlocale()[0] in ("zh_CN", "it_IT"),
                    reason="fail on a CI build with LC_ALL=zh_CN.utf8/it_IT.utf8",
                    strict=False,
                ),
            ),
            pytest.param(
                "01/10/2010 09:12:56 AM",
                "%m/%d/%Y %I:%M:%S %p",
                Timestamp("2010-01-10 09:12:56"),
                marks=pytest.mark.xfail(
                    locale.getlocale()[0] in ("zh_CN", "it_IT"),
                    reason="fail on a CI build with LC_ALL=zh_CN.utf8/it_IT.utf8",
                    strict=False,
                ),
            ),
        ],
    )
    def test_to_datetime_format_time(self, cache, value, format, dt):
        assert to_datetime(value, format=format, cache=cache) == dt

    @td.skip_if_not_us_locale
    def test_to_datetime_with_non_exact(self, cache):
        # GH 10834
        # 8904
        # exact kw
        ser = Series(
            ["19MAY11", "foobar19MAY11", "19MAY11:00:00:00", "19MAY11 00:00:00Z"]
        )
        result = to_datetime(ser, format="%d%b%y", exact=False, cache=cache)
        expected = to_datetime(
            ser.str.extract(r"(\d+\w+\d+)", expand=False), format="%d%b%y", cache=cache
        )
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize(
        "format, expected",
        [
            ("%Y-%m-%d", Timestamp(2000, 1, 3)),
            ("%Y-%d-%m", Timestamp(2000, 3, 1)),
            ("%Y-%m-%d %H", Timestamp(2000, 1, 3, 12)),
            ("%Y-%d-%m %H", Timestamp(2000, 3, 1, 12)),
            ("%Y-%m-%d %H:%M", Timestamp(2000, 1, 3, 12, 34)),
            ("%Y-%d-%m %H:%M", Timestamp(2000, 3, 1, 12, 34)),
            ("%Y-%m-%d %H:%M:%S", Timestamp(2000, 1, 3, 12, 34, 56)),
            ("%Y-%d-%m %H:%M:%S", Timestamp(2000, 3, 1, 12, 34, 56)),
            ("%Y-%m-%d %H:%M:%S.%f", Timestamp(2000, 1, 3, 12, 34, 56, 123456)),
            ("%Y-%d-%m %H:%M:%S.%f", Timestamp(2000, 3, 1, 12, 34, 56, 123456)),
            (
                "%Y-%m-%d %H:%M:%S.%f%z",
                Timestamp(2000, 1, 3, 12, 34, 56, 123456, tz="UTC+01:00"),
            ),
            (
                "%Y-%d-%m %H:%M:%S.%f%z",
                Timestamp(2000, 3, 1, 12, 34, 56, 123456, tz="UTC+01:00"),
            ),
        ],
    )
    def test_non_exact_doesnt_parse_whole_string(self, cache, format, expected):
        # https://github.com/pandas-dev/pandas/issues/50412
        # the formats alternate between ISO8601 and non-ISO8601 to check both paths
        result = to_datetime(
            "2000-01-03 12:34:56.123456+01:00", format=format, exact=False
        )
        assert result == expected

    @pytest.mark.parametrize(
        "arg",
        [
            "2012-01-01 09:00:00.000000001",
            "2012-01-01 09:00:00.000001",
            "2012-01-01 09:00:00.001",
            "2012-01-01 09:00:00.001000",
            "2012-01-01 09:00:00.001000000",
        ],
    )
    def test_parse_nanoseconds_with_formula(self, cache, arg):
        # GH8989
        # truncating the nanoseconds when a format was provided
        expected = to_datetime(arg, cache=cache)
        result = to_datetime(arg, format="%Y-%m-%d %H:%M:%S.%f", cache=cache)
        assert result == expected

    @pytest.mark.parametrize(
        "value,fmt,expected",
        [
            ["2009324", "%Y%W%w", Timestamp("2009-08-13")],
            ["2013020", "%Y%U%w", Timestamp("2013-01-13")],
        ],
    )
    def test_to_datetime_format_weeks(self, value, fmt, expected, cache):
        assert to_datetime(value, format=fmt, cache=cache) == expected

    @pytest.mark.parametrize(
        "fmt,dates,expected_dates",
        [
            [
                "%Y-%m-%d %H:%M:%S %Z",
                ["2010-01-01 12:00:00 UTC"] * 2,
                [Timestamp("2010-01-01 12:00:00", tz="UTC")] * 2,
            ],
            [
                "%Y-%m-%d %H:%M:%S%z",
                ["2010-01-01 12:00:00+0100"] * 2,
                [
                    Timestamp(
                        "2010-01-01 12:00:00", tzinfo=timezone(timedelta(minutes=60))
                    )
                ]
                * 2,
            ],
            [
                "%Y-%m-%d %H:%M:%S %z",
                ["2010-01-01 12:00:00 +0100"] * 2,
                [
                    Timestamp(
                        "2010-01-01 12:00:00", tzinfo=timezone(timedelta(minutes=60))
                    )
                ]
                * 2,
            ],
            [
                "%Y-%m-%d %H:%M:%S %z",
                ["2010-01-01 12:00:00 Z", "2010-01-01 12:00:00 Z"],
                [
                    Timestamp(
                        "2010-01-01 12:00:00", tzinfo=pytz.FixedOffset(0)
                    ),  # pytz coerces to UTC
                    Timestamp("2010-01-01 12:00:00", tzinfo=pytz.FixedOffset(0)),
                ],
            ],
        ],
    )
    def test_to_datetime_parse_tzname_or_tzoffset(self, fmt, dates, expected_dates):
        # GH 13486
        result = to_datetime(dates, format=fmt)
        expected = Index(expected_dates)
        tm.assert_equal(result, expected)

    @pytest.mark.parametrize(
        "fmt,dates,expected_dates",
        [
            [
                "%Y-%m-%d %H:%M:%S %Z",
                [
                    "2010-01-01 12:00:00 UTC",
                    "2010-01-01 12:00:00 GMT",
                    "2010-01-01 12:00:00 US/Pacific",
                ],
                [
                    Timestamp("2010-01-01 12:00:00", tz="UTC"),
                    Timestamp("2010-01-01 12:00:00", tz="GMT"),
                    Timestamp("2010-01-01 12:00:00", tz="US/Pacific"),
                ],
            ],
            [
                "%Y-%m-%d %H:%M:%S %z",
                ["2010-01-01 12:00:00 +0100", "2010-01-01 12:00:00 -0100"],
                [
                    Timestamp(
                        "2010-01-01 12:00:00", tzinfo=timezone(timedelta(minutes=60))
                    ),
                    Timestamp(
                        "2010-01-01 12:00:00", tzinfo=timezone(timedelta(minutes=-60))
                    ),
                ],
            ],
        ],
    )
    def test_to_datetime_parse_tzname_or_tzoffset_utc_false_deprecated(
        self, fmt, dates, expected_dates
    ):
        # GH 13486, 50887
        msg = "parsing datetimes with mixed time zones will raise an error"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            result = to_datetime(dates, format=fmt)
        expected = Index(expected_dates)
        tm.assert_equal(result, expected)

    def test_to_datetime_parse_tzname_or_tzoffset_different_tz_to_utc(self):
        # GH 32792
        dates = [
            "2010-01-01 12:00:00 +0100",
            "2010-01-01 12:00:00 -0100",
            "2010-01-01 12:00:00 +0300",
            "2010-01-01 12:00:00 +0400",
        ]
        expected_dates = [
            "2010-01-01 11:00:00+00:00",
            "2010-01-01 13:00:00+00:00",
            "2010-01-01 09:00:00+00:00",
            "2010-01-01 08:00:00+00:00",
        ]
        fmt = "%Y-%m-%d %H:%M:%S %z"

        result = to_datetime(dates, format=fmt, utc=True)
        expected = DatetimeIndex(expected_dates)
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "offset", ["+0", "-1foo", "UTCbar", ":10", "+01:000:01", ""]
    )
    def test_to_datetime_parse_timezone_malformed(self, offset):
        fmt = "%Y-%m-%d %H:%M:%S %z"
        date = "2010-01-01 12:00:00 " + offset

        msg = "|".join(
            [
                r'^time data ".*" doesn\'t match format ".*", at position 0. '
                f"{PARSING_ERR_MSG}$",
                r'^unconverted data remains when parsing with format ".*": ".*", '
                f"at position 0. {PARSING_ERR_MSG}$",
            ]
        )
        with pytest.raises(ValueError, match=msg):
            to_datetime([date], format=fmt)

    def test_to_datetime_parse_timezone_keeps_name(self):
        # GH 21697
        fmt = "%Y-%m-%d %H:%M:%S %z"
        arg = Index(["2010-01-01 12:00:00 Z"], name="foo")
        result = to_datetime(arg, format=fmt)
        expected = DatetimeIndex(["2010-01-01 12:00:00"], tz="UTC", name="foo")
        tm.assert_index_equal(result, expected)


class TestToDatetime:
    @pytest.mark.filterwarnings("ignore:Could not infer format")
    def test_to_datetime_overflow(self):
        # we should get an OutOfBoundsDatetime, NOT OverflowError
        # TODO: Timestamp raises ValueError("could not convert string to Timestamp")
        #  can we make these more consistent?
        arg = "08335394550"
        msg = 'Parsing "08335394550" to datetime overflows, at position 0'
        with pytest.raises(OutOfBoundsDatetime, match=msg):
            to_datetime(arg)

        with pytest.raises(OutOfBoundsDatetime, match=msg):
            to_datetime([arg])

        res = to_datetime(arg, errors="coerce")
        assert res is NaT
        res = to_datetime([arg], errors="coerce")
        tm.assert_index_equal(res, Index([NaT]))

        res = to_datetime(arg, errors="ignore")
        assert isinstance(res, str) and res == arg
        res = to_datetime([arg], errors="ignore")
        tm.assert_index_equal(res, Index([arg], dtype=object))

    def test_to_datetime_mixed_datetime_and_string(self):
        # GH#47018 adapted old doctest with new behavior
        d1 = datetime(2020, 1, 1, 17, tzinfo=timezone(-timedelta(hours=1)))
        d2 = datetime(2020, 1, 1, 18, tzinfo=timezone(-timedelta(hours=1)))
        res = to_datetime(["2020-01-01 17:00 -0100", d2])
        expected = to_datetime([d1, d2]).tz_convert(timezone(timedelta(minutes=-60)))
        tm.assert_index_equal(res, expected)

    def test_to_datetime_mixed_string_and_numeric(self):
        # GH#55780 np.array(vals) would incorrectly cast the number to str
        vals = ["2016-01-01", 0]
        expected = DatetimeIndex([Timestamp(x) for x in vals])
        result = to_datetime(vals, format="mixed")
        result2 = to_datetime(vals[::-1], format="mixed")[::-1]
        result3 = DatetimeIndex(vals)
        result4 = DatetimeIndex(vals[::-1])[::-1]

        tm.assert_index_equal(result, expected)
        tm.assert_index_equal(result2, expected)
        tm.assert_index_equal(result3, expected)
        tm.assert_index_equal(result4, expected)

    @pytest.mark.parametrize(
        "format", ["%Y-%m-%d", "%Y-%d-%m"], ids=["ISO8601", "non-ISO8601"]
    )
    def test_to_datetime_mixed_date_and_string(self, format):
        # https://github.com/pandas-dev/pandas/issues/50108
        d1 = date(2020, 1, 2)
        res = to_datetime(["2020-01-01", d1], format=format)
        expected = DatetimeIndex(["2020-01-01", "2020-01-02"], dtype="M8[ns]")
        tm.assert_index_equal(res, expected)

    @pytest.mark.parametrize(
        "fmt",
        ["%Y-%d-%m %H:%M:%S%z", "%Y-%m-%d %H:%M:%S%z"],
        ids=["non-ISO8601 format", "ISO8601 format"],
    )
    @pytest.mark.parametrize(
        "utc, args, expected",
        [
            pytest.param(
                True,
                ["2000-01-01 01:00:00-08:00", "2000-01-01 02:00:00-08:00"],
                DatetimeIndex(
                    ["2000-01-01 09:00:00+00:00", "2000-01-01 10:00:00+00:00"],
                    dtype="datetime64[ns, UTC]",
                ),
                id="all tz-aware, with utc",
            ),
            pytest.param(
                False,
                ["2000-01-01 01:00:00+00:00", "2000-01-01 02:00:00+00:00"],
                DatetimeIndex(
                    ["2000-01-01 01:00:00+00:00", "2000-01-01 02:00:00+00:00"],
                ),
                id="all tz-aware, without utc",
            ),
            pytest.param(
                True,
                ["2000-01-01 01:00:00-08:00", "2000-01-01 02:00:00+00:00"],
                DatetimeIndex(
                    ["2000-01-01 09:00:00+00:00", "2000-01-01 02:00:00+00:00"],
                    dtype="datetime64[ns, UTC]",
                ),
                id="all tz-aware, mixed offsets, with utc",
            ),
            pytest.param(
                True,
                ["2000-01-01 01:00:00", "2000-01-01 02:00:00+00:00"],
                DatetimeIndex(
                    ["2000-01-01 01:00:00+00:00", "2000-01-01 02:00:00+00:00"],
                    dtype="datetime64[ns, UTC]",
                ),
                id="tz-aware string, naive pydatetime, with utc",
            ),
        ],
    )
    @pytest.mark.parametrize(
        "constructor",
        [Timestamp, lambda x: Timestamp(x).to_pydatetime()],
    )
    def test_to_datetime_mixed_datetime_and_string_with_format(
        self, fmt, utc, args, expected, constructor
    ):
        # https://github.com/pandas-dev/pandas/issues/49298
        # https://github.com/pandas-dev/pandas/issues/50254
        # note: ISO8601 formats go down a fastpath, so we need to check both
        # a ISO8601 format and a non-ISO8601 one
        ts1 = constructor(args[0])
        ts2 = args[1]
        result = to_datetime([ts1, ts2], format=fmt, utc=utc)
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "fmt",
        ["%Y-%d-%m %H:%M:%S%z", "%Y-%m-%d %H:%M:%S%z"],
        ids=["non-ISO8601 format", "ISO8601 format"],
    )
    @pytest.mark.parametrize(
        "constructor",
        [Timestamp, lambda x: Timestamp(x).to_pydatetime()],
    )
    def test_to_datetime_mixed_datetime_and_string_with_format_mixed_offsets_utc_false(
        self, fmt, constructor
    ):
        # https://github.com/pandas-dev/pandas/issues/49298
        # https://github.com/pandas-dev/pandas/issues/50254
        # note: ISO8601 formats go down a fastpath, so we need to check both
        # a ISO8601 format and a non-ISO8601 one
        args = ["2000-01-01 01:00:00", "2000-01-01 02:00:00+00:00"]
        ts1 = constructor(args[0])
        ts2 = args[1]
        msg = "parsing datetimes with mixed time zones will raise an error"

        expected = Index(
            [
                Timestamp("2000-01-01 01:00:00"),
                Timestamp("2000-01-01 02:00:00+0000", tz="UTC"),
            ],
        )
        with tm.assert_produces_warning(FutureWarning, match=msg):
            result = to_datetime([ts1, ts2], format=fmt, utc=False)
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "fmt, expected",
        [
            pytest.param(
                "%Y-%m-%d %H:%M:%S%z",
                Index(
                    [
                        Timestamp("2000-01-01 09:00:00+0100", tz="UTC+01:00"),
                        Timestamp("2000-01-02 02:00:00+0200", tz="UTC+02:00"),
                        NaT,
                    ]
                ),
                id="ISO8601, non-UTC",
            ),
            pytest.param(
                "%Y-%d-%m %H:%M:%S%z",
                Index(
                    [
                        Timestamp("2000-01-01 09:00:00+0100", tz="UTC+01:00"),
                        Timestamp("2000-02-01 02:00:00+0200", tz="UTC+02:00"),
                        NaT,
                    ]
                ),
                id="non-ISO8601, non-UTC",
            ),
        ],
    )
    def test_to_datetime_mixed_offsets_with_none_tz(self, fmt, expected):
        # https://github.com/pandas-dev/pandas/issues/50071
        msg = "parsing datetimes with mixed time zones will raise an error"

        with tm.assert_produces_warning(FutureWarning, match=msg):
            result = to_datetime(
                ["2000-01-01 09:00:00+01:00", "2000-01-02 02:00:00+02:00", None],
                format=fmt,
                utc=False,
            )
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "fmt, expected",
        [
            pytest.param(
                "%Y-%m-%d %H:%M:%S%z",
                DatetimeIndex(
                    ["2000-01-01 08:00:00+00:00", "2000-01-02 00:00:00+00:00", "NaT"],
                    dtype="datetime64[ns, UTC]",
                ),
                id="ISO8601, UTC",
            ),
            pytest.param(
                "%Y-%d-%m %H:%M:%S%z",
                DatetimeIndex(
                    ["2000-01-01 08:00:00+00:00", "2000-02-01 00:00:00+00:00", "NaT"],
                    dtype="datetime64[ns, UTC]",
                ),
                id="non-ISO8601, UTC",
            ),
        ],
    )
    def test_to_datetime_mixed_offsets_with_none(self, fmt, expected):
        # https://github.com/pandas-dev/pandas/issues/50071
        result = to_datetime(
            ["2000-01-01 09:00:00+01:00", "2000-01-02 02:00:00+02:00", None],
            format=fmt,
            utc=True,
        )
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "fmt",
        ["%Y-%d-%m %H:%M:%S%z", "%Y-%m-%d %H:%M:%S%z"],
        ids=["non-ISO8601 format", "ISO8601 format"],
    )
    @pytest.mark.parametrize(
        "args",
        [
            pytest.param(
                ["2000-01-01 01:00:00-08:00", "2000-01-01 02:00:00-07:00"],
                id="all tz-aware, mixed timezones, without utc",
            ),
        ],
    )
    @pytest.mark.parametrize(
        "constructor",
        [Timestamp, lambda x: Timestamp(x).to_pydatetime()],
    )
    def test_to_datetime_mixed_datetime_and_string_with_format_raises(
        self, fmt, args, constructor
    ):
        # https://github.com/pandas-dev/pandas/issues/49298
        # note: ISO8601 formats go down a fastpath, so we need to check both
        # a ISO8601 format and a non-ISO8601 one
        ts1 = constructor(args[0])
        ts2 = constructor(args[1])
        with pytest.raises(
            ValueError, match="cannot be converted to datetime64 unless utc=True"
        ):
            to_datetime([ts1, ts2], format=fmt, utc=False)

    def test_to_datetime_np_str(self):
        # GH#32264
        # GH#48969
        value = np.str_("2019-02-04 10:18:46.297000+0000")

        ser = Series([value])

        exp = Timestamp("2019-02-04 10:18:46.297000", tz="UTC")

        assert to_datetime(value) == exp
        assert to_datetime(ser.iloc[0]) == exp

        res = to_datetime([value])
        expected = Index([exp])
        tm.assert_index_equal(res, expected)

        res = to_datetime(ser)
        expected = Series(expected)
        tm.assert_series_equal(res, expected)

    @pytest.mark.parametrize(
        "s, _format, dt",
        [
            ["2015-1-1", "%G-%V-%u", datetime(2014, 12, 29, 0, 0)],
            ["2015-1-4", "%G-%V-%u", datetime(2015, 1, 1, 0, 0)],
            ["2015-1-7", "%G-%V-%u", datetime(2015, 1, 4, 0, 0)],
        ],
    )
    def test_to_datetime_iso_week_year_format(self, s, _format, dt):
        # See GH#16607
        assert to_datetime(s, format=_format) == dt

    @pytest.mark.parametrize(
        "msg, s, _format",
        [
            [
                "ISO week directive '%V' is incompatible with the year directive "
                "'%Y'. Use the ISO year '%G' instead.",
                "1999 50",
                "%Y %V",
            ],
            [
                "ISO year directive '%G' must be used with the ISO week directive "
                "'%V' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "1999 51",
                "%G %V",
            ],
            [
                "ISO year directive '%G' must be used with the ISO week directive "
                "'%V' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "1999 Monday",
                "%G %A",
            ],
            [
                "ISO year directive '%G' must be used with the ISO week directive "
                "'%V' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "1999 Mon",
                "%G %a",
            ],
            [
                "ISO year directive '%G' must be used with the ISO week directive "
                "'%V' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "1999 6",
                "%G %w",
            ],
            [
                "ISO year directive '%G' must be used with the ISO week directive "
                "'%V' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "1999 6",
                "%G %u",
            ],
            [
                "ISO year directive '%G' must be used with the ISO week directive "
                "'%V' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "2051",
                "%G",
            ],
            [
                "Day of the year directive '%j' is not compatible with ISO year "
                "directive '%G'. Use '%Y' instead.",
                "1999 51 6 256",
                "%G %V %u %j",
            ],
            [
                "ISO week directive '%V' is incompatible with the year directive "
                "'%Y'. Use the ISO year '%G' instead.",
                "1999 51 Sunday",
                "%Y %V %A",
            ],
            [
                "ISO week directive '%V' is incompatible with the year directive "
                "'%Y'. Use the ISO year '%G' instead.",
                "1999 51 Sun",
                "%Y %V %a",
            ],
            [
                "ISO week directive '%V' is incompatible with the year directive "
                "'%Y'. Use the ISO year '%G' instead.",
                "1999 51 1",
                "%Y %V %w",
            ],
            [
                "ISO week directive '%V' is incompatible with the year directive "
                "'%Y'. Use the ISO year '%G' instead.",
                "1999 51 1",
                "%Y %V %u",
            ],
            [
                "ISO week directive '%V' must be used with the ISO year directive "
                "'%G' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "20",
                "%V",
            ],
            [
                "ISO week directive '%V' must be used with the ISO year directive "
                "'%G' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "1999 51 Sunday",
                "%V %A",
            ],
            [
                "ISO week directive '%V' must be used with the ISO year directive "
                "'%G' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "1999 51 Sun",
                "%V %a",
            ],
            [
                "ISO week directive '%V' must be used with the ISO year directive "
                "'%G' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "1999 51 1",
                "%V %w",
            ],
            [
                "ISO week directive '%V' must be used with the ISO year directive "
                "'%G' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "1999 51 1",
                "%V %u",
            ],
            [
                "Day of the year directive '%j' is not compatible with ISO year "
                "directive '%G'. Use '%Y' instead.",
                "1999 50",
                "%G %j",
            ],
            [
                "ISO week directive '%V' must be used with the ISO year directive "
                "'%G' and a weekday directive '%A', '%a', '%w', or '%u'.",
                "20 Monday",
                "%V %A",
            ],
        ],
    )
    @pytest.mark.parametrize("errors", ["raise", "coerce", "ignore"])
    def test_error_iso_week_year(self, msg, s, _format, errors):
        # See GH#16607, GH#50308
        # This test checks for errors thrown when giving the wrong format
        # However, as discussed on PR#25541, overriding the locale
        # causes a different error to be thrown due to the format being
        # locale specific, but the test data is in english.
        # Therefore, the tests only run when locale is not overwritten,
        # as a sort of solution to this problem.
        if locale.getlocale() != ("zh_CN", "UTF-8") and locale.getlocale() != (
            "it_IT",
            "UTF-8",
        ):
            with pytest.raises(ValueError, match=msg):
                to_datetime(s, format=_format, errors=errors)

    @pytest.mark.parametrize("tz", [None, "US/Central"])
    def test_to_datetime_dtarr(self, tz):
        # DatetimeArray
        dti = date_range("1965-04-03", periods=19, freq="2W", tz=tz)
        arr = dti._data

        result = to_datetime(arr)
        assert result is arr

    # Doesn't work on Windows since tzpath not set correctly
    @td.skip_if_windows
    @pytest.mark.parametrize("arg_class", [Series, Index])
    @pytest.mark.parametrize("utc", [True, False])
    @pytest.mark.parametrize("tz", [None, "US/Central"])
    def test_to_datetime_arrow(self, tz, utc, arg_class):
        pa = pytest.importorskip("pyarrow")

        dti = date_range("1965-04-03", periods=19, freq="2W", tz=tz)
        dti = arg_class(dti)

        dti_arrow = dti.astype(pd.ArrowDtype(pa.timestamp(unit="ns", tz=tz)))

        result = to_datetime(dti_arrow, utc=utc)
        expected = to_datetime(dti, utc=utc).astype(
            pd.ArrowDtype(pa.timestamp(unit="ns", tz=tz if not utc else "UTC"))
        )
        if not utc and arg_class is not Series:
            # Doesn't hold for utc=True, since that will astype
            # to_datetime also returns a new object for series
            assert result is dti_arrow
        if arg_class is Series:
            tm.assert_series_equal(result, expected)
        else:
            tm.assert_index_equal(result, expected)

    def test_to_datetime_pydatetime(self):
        actual = to_datetime(datetime(2008, 1, 15))
        assert actual == datetime(2008, 1, 15)

    def test_to_datetime_YYYYMMDD(self):
        actual = to_datetime("20080115")
        assert actual == datetime(2008, 1, 15)

    def test_to_datetime_unparsable_ignore(self):
        # unparsable
        ser = "Month 1, 1999"
        assert to_datetime(ser, errors="ignore") == ser

    @td.skip_if_windows  # `tm.set_timezone` does not work in windows
    def test_to_datetime_now(self):
        # See GH#18666
        with tm.set_timezone("US/Eastern"):
            # GH#18705
            now = Timestamp("now").as_unit("ns")
            pdnow = to_datetime("now")
            pdnow2 = to_datetime(["now"])[0]

            # These should all be equal with infinite perf; this gives
            # a generous margin of 10 seconds
            assert abs(pdnow._value - now._value) < 1e10
            assert abs(pdnow2._value - now._value) < 1e10

            assert pdnow.tzinfo is None
            assert pdnow2.tzinfo is None

    @td.skip_if_windows  # `tm.set_timezone` does not work in windows
    @pytest.mark.parametrize("tz", ["Pacific/Auckland", "US/Samoa"])
    def test_to_datetime_today(self, tz):
        # See GH#18666
        # Test with one timezone far ahead of UTC and another far behind, so
        # one of these will _almost_ always be in a different day from UTC.
        # Unfortunately this test between 12 and 1 AM Samoa time
        # this both of these timezones _and_ UTC will all be in the same day,
        # so this test will not detect the regression introduced in #18666.
        with tm.set_timezone(tz):
            nptoday = np.datetime64("today").astype("datetime64[ns]").astype(np.int64)
            pdtoday = to_datetime("today")
            pdtoday2 = to_datetime(["today"])[0]

            tstoday = Timestamp("today").as_unit("ns")
            tstoday2 = Timestamp.today().as_unit("ns")

            # These should all be equal with infinite perf; this gives
            # a generous margin of 10 seconds
            assert abs(pdtoday.normalize()._value - nptoday) < 1e10
            assert abs(pdtoday2.normalize()._value - nptoday) < 1e10
            assert abs(pdtoday._value - tstoday._value) < 1e10
            assert abs(pdtoday._value - tstoday2._value) < 1e10

            assert pdtoday.tzinfo is None
            assert pdtoday2.tzinfo is None

    @pytest.mark.parametrize("arg", ["now", "today"])
    def test_to_datetime_today_now_unicode_bytes(self, arg):
        to_datetime([arg])

    @pytest.mark.parametrize(
        "format, expected_ds",
        [
            ("%Y-%m-%d %H:%M:%S%z", "2020-01-03"),
            ("%Y-%d-%m %H:%M:%S%z", "2020-03-01"),
            (None, "2020-01-03"),
        ],
    )
    @pytest.mark.parametrize(
        "string, attribute",
        [
            ("now", "utcnow"),
            ("today", "today"),
        ],
    )
    def test_to_datetime_now_with_format(self, format, expected_ds, string, attribute):
        # https://github.com/pandas-dev/pandas/issues/50359
        result = to_datetime(["2020-01-03 00:00:00Z", string], format=format, utc=True)
        expected = DatetimeIndex(
            [expected_ds, getattr(Timestamp, attribute)()], dtype="datetime64[ns, UTC]"
        )
        assert (expected - result).max().total_seconds() < 1

    @pytest.mark.parametrize(
        "dt", [np.datetime64("2000-01-01"), np.datetime64("2000-01-02")]
    )
    def test_to_datetime_dt64s(self, cache, dt):
        assert to_datetime(dt, cache=cache) == Timestamp(dt)

    @pytest.mark.parametrize(
        "arg, format",
        [
            ("2001-01-01", "%Y-%m-%d"),
            ("01-01-2001", "%d-%m-%Y"),
        ],
    )
    def test_to_datetime_dt64s_and_str(self, arg, format):
        # https://github.com/pandas-dev/pandas/issues/50036
        result = to_datetime([arg, np.datetime64("2020-01-01")], format=format)
        expected = DatetimeIndex(["2001-01-01", "2020-01-01"])
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "dt", [np.datetime64("1000-01-01"), np.datetime64("5000-01-02")]
    )
    @pytest.mark.parametrize("errors", ["raise", "ignore", "coerce"])
    def test_to_datetime_dt64s_out_of_ns_bounds(self, cache, dt, errors):
        # GH#50369 We cast to the nearest supported reso, i.e. "s"
        ts = to_datetime(dt, errors=errors, cache=cache)
        assert isinstance(ts, Timestamp)
        assert ts.unit == "s"
        assert ts.asm8 == dt

        ts = Timestamp(dt)
        assert ts.unit == "s"
        assert ts.asm8 == dt

    @pytest.mark.skip_ubsan
    def test_to_datetime_dt64d_out_of_bounds(self, cache):
        dt64 = np.datetime64(np.iinfo(np.int64).max, "D")

        msg = "Out of bounds second timestamp: 25252734927768524-07-27"
        with pytest.raises(OutOfBoundsDatetime, match=msg):
            Timestamp(dt64)
        with pytest.raises(OutOfBoundsDatetime, match=msg):
            to_datetime(dt64, errors="raise", cache=cache)

        assert to_datetime(dt64, errors="coerce", cache=cache) is NaT

    @pytest.mark.parametrize("unit", ["s", "D"])
    def test_to_datetime_array_of_dt64s(self, cache, unit):
        # https://github.com/pandas-dev/pandas/issues/31491
        # Need at least 50 to ensure cache is used.
        dts = [
            np.datetime64("2000-01-01", unit),
            np.datetime64("2000-01-02", unit),
        ] * 30
        # Assuming all datetimes are in bounds, to_datetime() returns
        # an array that is equal to Timestamp() parsing
        result = to_datetime(dts, cache=cache)
        if cache:
            # FIXME: behavior should not depend on cache
            expected = DatetimeIndex([Timestamp(x).asm8 for x in dts], dtype="M8[s]")
        else:
            expected = DatetimeIndex([Timestamp(x).asm8 for x in dts], dtype="M8[ns]")

        tm.assert_index_equal(result, expected)

        # A list of datetimes where the last one is out of bounds
        dts_with_oob = dts + [np.datetime64("9999-01-01")]

        # As of GH#51978 we do not raise in this case
        to_datetime(dts_with_oob, errors="raise")

        result = to_datetime(dts_with_oob, errors="coerce", cache=cache)
        if not cache:
            # FIXME: shouldn't depend on cache!
            expected = DatetimeIndex(
                [Timestamp(dts_with_oob[0]).asm8, Timestamp(dts_with_oob[1]).asm8] * 30
                + [NaT],
            )
        else:
            expected = DatetimeIndex(np.array(dts_with_oob, dtype="M8[s]"))
        tm.assert_index_equal(result, expected)

        # With errors='ignore', out of bounds datetime64s
        # are converted to their .item(), which depending on the version of
        # numpy is either a python datetime.datetime or datetime.date
        result = to_datetime(dts_with_oob, errors="ignore", cache=cache)
        if not cache:
            # FIXME: shouldn't depend on cache!
            expected = Index(dts_with_oob)
        tm.assert_index_equal(result, expected)

    def test_out_of_bounds_errors_ignore(self):
        # https://github.com/pandas-dev/pandas/issues/50587
        result = to_datetime(np.datetime64("9999-01-01"), errors="ignore")
        expected = np.datetime64("9999-01-01")
        assert result == expected

    def test_out_of_bounds_errors_ignore2(self):
        # GH#12424
        msg = "errors='ignore' is deprecated"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            res = to_datetime(Series(["2362-01-01", np.nan]), errors="ignore")
        exp = Series(["2362-01-01", np.nan])
        tm.assert_series_equal(res, exp)

    def test_to_datetime_tz(self, cache):
        # xref 8260
        # uniform returns a DatetimeIndex
        arr = [
            Timestamp("2013-01-01 13:00:00-0800", tz="US/Pacific"),
            Timestamp("2013-01-02 14:00:00-0800", tz="US/Pacific"),
        ]
        result = to_datetime(arr, cache=cache)
        expected = DatetimeIndex(
            ["2013-01-01 13:00:00", "2013-01-02 14:00:00"], tz="US/Pacific"
        )
        tm.assert_index_equal(result, expected)

    def test_to_datetime_tz_mixed(self, cache):
        # mixed tzs will raise if errors='raise'
        # https://github.com/pandas-dev/pandas/issues/50585
        arr = [
            Timestamp("2013-01-01 13:00:00", tz="US/Pacific"),
            Timestamp("2013-01-02 14:00:00", tz="US/Eastern"),
        ]
        msg = (
            "Tz-aware datetime.datetime cannot be "
            "converted to datetime64 unless utc=True"
        )
        with pytest.raises(ValueError, match=msg):
            to_datetime(arr, cache=cache)

        depr_msg = "errors='ignore' is deprecated"
        with tm.assert_produces_warning(FutureWarning, match=depr_msg):
            result = to_datetime(arr, cache=cache, errors="ignore")
        expected = Index(
            [
                Timestamp("2013-01-01 13:00:00-08:00"),
                Timestamp("2013-01-02 14:00:00-05:00"),
            ],
            dtype="object",
        )
        tm.assert_index_equal(result, expected)
        result = to_datetime(arr, cache=cache, errors="coerce")
        expected = DatetimeIndex(
            ["2013-01-01 13:00:00-08:00", "NaT"], dtype="datetime64[ns, US/Pacific]"
        )
        tm.assert_index_equal(result, expected)

    def test_to_datetime_different_offsets(self, cache):
        # inspired by asv timeseries.ToDatetimeNONISO8601 benchmark
        # see GH-26097 for more
        ts_string_1 = "March 1, 2018 12:00:00+0400"
        ts_string_2 = "March 1, 2018 12:00:00+0500"
        arr = [ts_string_1] * 5 + [ts_string_2] * 5
        expected = Index([parse(x) for x in arr])
        msg = "parsing datetimes with mixed time zones will raise an error"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            result = to_datetime(arr, cache=cache)
        tm.assert_index_equal(result, expected)

    def test_to_datetime_tz_pytz(self, cache):
        # see gh-8260
        us_eastern = pytz.timezone("US/Eastern")
        arr = np.array(
            [
                us_eastern.localize(
                    datetime(year=2000, month=1, day=1, hour=3, minute=0)
                ),
                us_eastern.localize(
                    datetime(year=2000, month=6, day=1, hour=3, minute=0)
                ),
            ],
            dtype=object,
        )
        result = to_datetime(arr, utc=True, cache=cache)
        expected = DatetimeIndex(
            ["2000-01-01 08:00:00+00:00", "2000-06-01 07:00:00+00:00"],
            dtype="datetime64[ns, UTC]",
            freq=None,
        )
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "init_constructor, end_constructor",
        [
            (Index, DatetimeIndex),
            (list, DatetimeIndex),
            (np.array, DatetimeIndex),
            (Series, Series),
        ],
    )
    def test_to_datetime_utc_true(self, cache, init_constructor, end_constructor):
        # See gh-11934 & gh-6415
        data = ["20100102 121314", "20100102 121315"]
        expected_data = [
            Timestamp("2010-01-02 12:13:14", tz="utc"),
            Timestamp("2010-01-02 12:13:15", tz="utc"),
        ]

        result = to_datetime(
            init_constructor(data), format="%Y%m%d %H%M%S", utc=True, cache=cache
        )
        expected = end_constructor(expected_data)
        tm.assert_equal(result, expected)

    @pytest.mark.parametrize(
        "scalar, expected",
        [
            ["20100102 121314", Timestamp("2010-01-02 12:13:14", tz="utc")],
            ["20100102 121315", Timestamp("2010-01-02 12:13:15", tz="utc")],
        ],
    )
    def test_to_datetime_utc_true_scalar(self, cache, scalar, expected):
        # Test scalar case as well
        result = to_datetime(scalar, format="%Y%m%d %H%M%S", utc=True, cache=cache)
        assert result == expected

    def test_to_datetime_utc_true_with_series_single_value(self, cache):
        # GH 15760 UTC=True with Series
        ts = 1.5e18
        result = to_datetime(Series([ts]), utc=True, cache=cache)
        expected = Series([Timestamp(ts, tz="utc")])
        tm.assert_series_equal(result, expected)

    def test_to_datetime_utc_true_with_series_tzaware_string(self, cache):
        ts = "2013-01-01 00:00:00-01:00"
        expected_ts = "2013-01-01 01:00:00"
        data = Series([ts] * 3)
        result = to_datetime(data, utc=True, cache=cache)
        expected = Series([Timestamp(expected_ts, tz="utc")] * 3)
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize(
        "date, dtype",
        [
            ("2013-01-01 01:00:00", "datetime64[ns]"),
            ("2013-01-01 01:00:00", "datetime64[ns, UTC]"),
        ],
    )
    def test_to_datetime_utc_true_with_series_datetime_ns(self, cache, date, dtype):
        expected = Series(
            [Timestamp("2013-01-01 01:00:00", tz="UTC")], dtype="M8[ns, UTC]"
        )
        result = to_datetime(Series([date], dtype=dtype), utc=True, cache=cache)
        tm.assert_series_equal(result, expected)

    def test_to_datetime_tz_psycopg2(self, request, cache):
        # xref 8260
        psycopg2_tz = pytest.importorskip("psycopg2.tz")

        # misc cases
        tz1 = psycopg2_tz.FixedOffsetTimezone(offset=-300, name=None)
        tz2 = psycopg2_tz.FixedOffsetTimezone(offset=-240, name=None)
        arr = np.array(
            [
                datetime(2000, 1, 1, 3, 0, tzinfo=tz1),
                datetime(2000, 6, 1, 3, 0, tzinfo=tz2),
            ],
            dtype=object,
        )

        result = to_datetime(arr, errors="coerce", utc=True, cache=cache)
        expected = DatetimeIndex(
            ["2000-01-01 08:00:00+00:00", "2000-06-01 07:00:00+00:00"],
            dtype="datetime64[ns, UTC]",
            freq=None,
        )
        tm.assert_index_equal(result, expected)

        # dtype coercion
        i = DatetimeIndex(
            ["2000-01-01 08:00:00"],
            tz=psycopg2_tz.FixedOffsetTimezone(offset=-300, name=None),
        )
        assert is_datetime64_ns_dtype(i)

        # tz coercion
        result = to_datetime(i, errors="coerce", cache=cache)
        tm.assert_index_equal(result, i)

        result = to_datetime(i, errors="coerce", utc=True, cache=cache)
        expected = DatetimeIndex(["2000-01-01 13:00:00"], dtype="datetime64[ns, UTC]")
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize("arg", [True, False])
    def test_datetime_bool(self, cache, arg):
        # GH13176
        msg = r"dtype bool cannot be converted to datetime64\[ns\]"
        with pytest.raises(TypeError, match=msg):
            to_datetime(arg)
        assert to_datetime(arg, errors="coerce", cache=cache) is NaT
        assert to_datetime(arg, errors="ignore", cache=cache) is arg

    def test_datetime_bool_arrays_mixed(self, cache):
        msg = f"{type(cache)} is not convertible to datetime"
        with pytest.raises(TypeError, match=msg):
            to_datetime([False, datetime.today()], cache=cache)
        with pytest.raises(
            ValueError,
            match=(
                r'^time data "True" doesn\'t match format "%Y%m%d", '
                f"at position 1. {PARSING_ERR_MSG}$"
            ),
        ):
            to_datetime(["20130101", True], cache=cache)
        tm.assert_index_equal(
            to_datetime([0, False, NaT, 0.0], errors="coerce", cache=cache),
            DatetimeIndex(
                [to_datetime(0, cache=cache), NaT, NaT, to_datetime(0, cache=cache)]
            ),
        )

    @pytest.mark.parametrize("arg", [bool, to_datetime])
    def test_datetime_invalid_datatype(self, arg):
        # GH13176
        msg = "is not convertible to datetime"
        with pytest.raises(TypeError, match=msg):
            to_datetime(arg)

    @pytest.mark.parametrize("errors", ["coerce", "raise", "ignore"])
    def test_invalid_format_raises(self, errors):
        # https://github.com/pandas-dev/pandas/issues/50255
        with pytest.raises(
            ValueError, match="':' is a bad directive in format 'H%:M%:S%"
        ):
            to_datetime(["00:00:00"], format="H%:M%:S%", errors=errors)

    @pytest.mark.parametrize("value", ["a", "00:01:99"])
    @pytest.mark.parametrize("format", [None, "%H:%M:%S"])
    def test_datetime_invalid_scalar(self, value, format):
        # GH24763
        res = to_datetime(value, errors="ignore", format=format)
        assert res == value

        res = to_datetime(value, errors="coerce", format=format)
        assert res is NaT

        msg = "|".join(
            [
                r'^time data "a" doesn\'t match format "%H:%M:%S", at position 0. '
                f"{PARSING_ERR_MSG}$",
                r'^Given date string "a" not likely a datetime, at position 0$',
                r'^unconverted data remains when parsing with format "%H:%M:%S": "9", '
                f"at position 0. {PARSING_ERR_MSG}$",
                r"^second must be in 0..59: 00:01:99, at position 0$",
            ]
        )
        with pytest.raises(ValueError, match=msg):
            to_datetime(value, errors="raise", format=format)

    @pytest.mark.parametrize("value", ["3000/12/11 00:00:00"])
    @pytest.mark.parametrize("format", [None, "%H:%M:%S"])
    def test_datetime_outofbounds_scalar(self, value, format):
        # GH24763
        res = to_datetime(value, errors="ignore", format=format)
        assert res == value

        res = to_datetime(value, errors="coerce", format=format)
        assert res is NaT

        if format is not None:
            msg = r'^time data ".*" doesn\'t match format ".*", at position 0.'
            with pytest.raises(ValueError, match=msg):
                to_datetime(value, errors="raise", format=format)
        else:
            msg = "^Out of bounds .*, at position 0$"
            with pytest.raises(OutOfBoundsDatetime, match=msg):
                to_datetime(value, errors="raise", format=format)

    @pytest.mark.parametrize(
        ("values"), [(["a"]), (["00:01:99"]), (["a", "b", "99:00:00"])]
    )
    @pytest.mark.parametrize("format", [(None), ("%H:%M:%S")])
    def test_datetime_invalid_index(self, values, format):
        # GH24763
        # Not great to have logic in tests, but this one's hard to
        # parametrise over
        if format is None and len(values) > 1:
            warn = UserWarning
        else:
            warn = None
        with tm.assert_produces_warning(
            warn, match="Could not infer format", raise_on_extra_warnings=False
        ):
            res = to_datetime(values, errors="ignore", format=format)
        tm.assert_index_equal(
            res, Index(values, dtype="object" if format is None else "str")
        )

        with tm.assert_produces_warning(
            warn, match="Could not infer format", raise_on_extra_warnings=False
        ):
            res = to_datetime(values, errors="coerce", format=format)
        tm.assert_index_equal(res, DatetimeIndex([NaT] * len(values)))

        msg = "|".join(
            [
                r'^Given date string "a" not likely a datetime, at position 0$',
                r'^time data "a" doesn\'t match format "%H:%M:%S", at position 0. '
                f"{PARSING_ERR_MSG}$",
                r'^unconverted data remains when parsing with format "%H:%M:%S": "9", '
                f"at position 0. {PARSING_ERR_MSG}$",
                r"^second must be in 0..59: 00:01:99, at position 0$",
            ]
        )
        with pytest.raises(ValueError, match=msg):
            with tm.assert_produces_warning(
                warn, match="Could not infer format", raise_on_extra_warnings=False
            ):
                to_datetime(values, errors="raise", format=format)

    @pytest.mark.parametrize("utc", [True, None])
    @pytest.mark.parametrize("format", ["%Y%m%d %H:%M:%S", None])
    @pytest.mark.parametrize("constructor", [list, tuple, np.array, Index, deque])
    def test_to_datetime_cache(self, utc, format, constructor):
        date = "20130101 00:00:00"
        test_dates = [date] * 10**5
        data = constructor(test_dates)

        result = to_datetime(data, utc=utc, format=format, cache=True)
        expected = to_datetime(data, utc=utc, format=format, cache=False)

        tm.assert_index_equal(result, expected)

    def test_to_datetime_from_deque(self):
        # GH 29403
        result = to_datetime(deque([Timestamp("2010-06-02 09:30:00")] * 51))
        expected = to_datetime([Timestamp("2010-06-02 09:30:00")] * 51)
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize("utc", [True, None])
    @pytest.mark.parametrize("format", ["%Y%m%d %H:%M:%S", None])
    def test_to_datetime_cache_series(self, utc, format):
        date = "20130101 00:00:00"
        test_dates = [date] * 10**5
        data = Series(test_dates)
        result = to_datetime(data, utc=utc, format=format, cache=True)
        expected = to_datetime(data, utc=utc, format=format, cache=False)
        tm.assert_series_equal(result, expected)

    def test_to_datetime_cache_scalar(self):
        date = "20130101 00:00:00"
        result = to_datetime(date, cache=True)
        expected = Timestamp("20130101 00:00:00")
        assert result == expected

    @pytest.mark.parametrize(
        "datetimelikes,expected_values",
        (
            (
                (None, np.nan) + (NaT,) * start_caching_at,
                (NaT,) * (start_caching_at + 2),
            ),
            (
                (None, Timestamp("2012-07-26")) + (NaT,) * start_caching_at,
                (NaT, Timestamp("2012-07-26")) + (NaT,) * start_caching_at,
            ),
            (
                (None,)
                + (NaT,) * start_caching_at
                + ("2012 July 26", Timestamp("2012-07-26")),
                (NaT,) * (start_caching_at + 1)
                + (Timestamp("2012-07-26"), Timestamp("2012-07-26")),
            ),
        ),
    )
    def test_convert_object_to_datetime_with_cache(
        self, datetimelikes, expected_values
    ):
        # GH#39882
        ser = Series(
            datetimelikes,
            dtype="object",
        )
        result_series = to_datetime(ser, errors="coerce")
        expected_series = Series(
            expected_values,
            dtype="datetime64[ns]",
        )
        tm.assert_series_equal(result_series, expected_series)

    @pytest.mark.parametrize("cache", [True, False])
    @pytest.mark.parametrize(
        "input",
        [
            Series([NaT] * 20 + [None] * 20, dtype="object"),
            Series([NaT] * 60 + [None] * 60, dtype="object"),
            Series([None] * 20),
            Series([None] * 60),
            Series([""] * 20),
            Series([""] * 60),
            Series([pd.NA] * 20),
            Series([pd.NA] * 60),
            Series([np.nan] * 20),
            Series([np.nan] * 60),
        ],
    )
    def test_to_datetime_converts_null_like_to_nat(self, cache, input):
        # GH35888
        expected = Series([NaT] * len(input), dtype="M8[ns]")
        result = to_datetime(input, cache=cache)
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize(
        "date, format",
        [
            ("2017-20", "%Y-%W"),
            ("20 Sunday", "%W %A"),
            ("20 Sun", "%W %a"),
            ("2017-21", "%Y-%U"),
            ("20 Sunday", "%U %A"),
            ("20 Sun", "%U %a"),
        ],
    )
    def test_week_without_day_and_calendar_year(self, date, format):
        # GH16774

        msg = "Cannot use '%W' or '%U' without day and year"
        with pytest.raises(ValueError, match=msg):
            to_datetime(date, format=format)

    def test_to_datetime_coerce(self):
        # GH 26122
        ts_strings = [
            "March 1, 2018 12:00:00+0400",
            "March 1, 2018 12:00:00+0500",
            "20100240",
        ]
        msg = "parsing datetimes with mixed time zones will raise an error"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            result = to_datetime(ts_strings, errors="coerce")
        expected = Index(
            [
                datetime(2018, 3, 1, 12, 0, tzinfo=tzoffset(None, 14400)),
                datetime(2018, 3, 1, 12, 0, tzinfo=tzoffset(None, 18000)),
                NaT,
            ]
        )
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "string_arg, format",
        [("March 1, 2018", "%B %d, %Y"), ("2018-03-01", "%Y-%m-%d")],
    )
    @pytest.mark.parametrize(
        "outofbounds",
        [
            datetime(9999, 1, 1),
            date(9999, 1, 1),
            np.datetime64("9999-01-01"),
            "January 1, 9999",
            "9999-01-01",
        ],
    )
    def test_to_datetime_coerce_oob(self, string_arg, format, outofbounds):
        # https://github.com/pandas-dev/pandas/issues/50255
        ts_strings = [string_arg, outofbounds]
        result = to_datetime(ts_strings, errors="coerce", format=format)
        expected = DatetimeIndex([datetime(2018, 3, 1), NaT])
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "errors, expected",
        [
            ("coerce", Index([NaT, NaT])),
            ("ignore", Index(["200622-12-31", "111111-24-11"], dtype=object)),
        ],
    )
    def test_to_datetime_malformed_no_raise(self, errors, expected):
        # GH 28299
        # GH 48633
        ts_strings = ["200622-12-31", "111111-24-11"]
        with tm.assert_produces_warning(
            UserWarning, match="Could not infer format", raise_on_extra_warnings=False
        ):
            result = to_datetime(ts_strings, errors=errors)
        tm.assert_index_equal(result, expected)

    def test_to_datetime_malformed_raise(self):
        # GH 48633
        ts_strings = ["200622-12-31", "111111-24-11"]
        msg = (
            'Parsed string "200622-12-31" gives an invalid tzoffset, which must '
            r"be between -timedelta\(hours=24\) and timedelta\(hours=24\), "
            "at position 0"
        )
        with pytest.raises(
            ValueError,
            match=msg,
        ):
            with tm.assert_produces_warning(
                UserWarning, match="Could not infer format"
            ):
                to_datetime(
                    ts_strings,
                    errors="raise",
                )

    def test_iso_8601_strings_with_same_offset(self):
        # GH 17697, 11736
        ts_str = "2015-11-18 15:30:00+05:30"
        result = to_datetime(ts_str)
        expected = Timestamp(ts_str)
        assert result == expected

        expected = DatetimeIndex([Timestamp(ts_str)] * 2)
        result = to_datetime([ts_str] * 2)
        tm.assert_index_equal(result, expected)

        result = DatetimeIndex([ts_str] * 2)
        tm.assert_index_equal(result, expected)

    def test_iso_8601_strings_with_different_offsets(self):
        # GH 17697, 11736, 50887
        ts_strings = ["2015-11-18 15:30:00+05:30", "2015-11-18 16:30:00+06:30", NaT]
        msg = "parsing datetimes with mixed time zones will raise an error"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            result = to_datetime(ts_strings)
        expected = np.array(
            [
                datetime(2015, 11, 18, 15, 30, tzinfo=tzoffset(None, 19800)),
                datetime(2015, 11, 18, 16, 30, tzinfo=tzoffset(None, 23400)),
                NaT,
            ],
            dtype=object,
        )
        # GH 21864
        expected = Index(expected)
        tm.assert_index_equal(result, expected)

    def test_iso_8601_strings_with_different_offsets_utc(self):
        ts_strings = ["2015-11-18 15:30:00+05:30", "2015-11-18 16:30:00+06:30", NaT]
        result = to_datetime(ts_strings, utc=True)
        expected = DatetimeIndex(
            [Timestamp(2015, 11, 18, 10), Timestamp(2015, 11, 18, 10), NaT], tz="UTC"
        )
        tm.assert_index_equal(result, expected)

    def test_mixed_offsets_with_native_datetime_raises(self):
        # GH 25978

        vals = [
            "nan",
            Timestamp("1990-01-01"),
            "2015-03-14T16:15:14.123-08:00",
            "2019-03-04T21:56:32.620-07:00",
            None,
            "today",
            "now",
        ]
        ser = Series(vals)
        assert all(ser[i] is vals[i] for i in range(len(vals)))  # GH#40111

        now = Timestamp("now")
        today = Timestamp("today")
        msg = "parsing datetimes with mixed time zones will raise an error"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            mixed = to_datetime(ser)
        expected = Series(
            [
                "NaT",
                Timestamp("1990-01-01"),
                Timestamp("2015-03-14T16:15:14.123-08:00").to_pydatetime(),
                Timestamp("2019-03-04T21:56:32.620-07:00").to_pydatetime(),
                None,
            ],
            dtype=object,
        )
        tm.assert_series_equal(mixed[:-2], expected)
        # we'll check mixed[-1] and mixed[-2] match now and today to within
        # call-timing tolerances
        assert (now - mixed.iloc[-1]).total_seconds() <= 0.1
        assert (today - mixed.iloc[-2]).total_seconds() <= 0.1

        with pytest.raises(ValueError, match="Tz-aware datetime.datetime"):
            to_datetime(mixed)

    def test_non_iso_strings_with_tz_offset(self):
        result = to_datetime(["March 1, 2018 12:00:00+0400"] * 2)
        expected = DatetimeIndex(
            [datetime(2018, 3, 1, 12, tzinfo=timezone(timedelta(minutes=240)))] * 2
        )
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "ts, expected",
        [
            (Timestamp("2018-01-01"), Timestamp("2018-01-01", tz="UTC")),
            (
                Timestamp("2018-01-01", tz="US/Pacific"),
                Timestamp("2018-01-01 08:00", tz="UTC"),
            ),
        ],
    )
    def test_timestamp_utc_true(self, ts, expected):
        # GH 24415
        result = to_datetime(ts, utc=True)
        assert result == expected

    @pytest.mark.parametrize("dt_str", ["00010101", "13000101", "30000101", "99990101"])
    def test_to_datetime_with_format_out_of_bounds(self, dt_str):
        # GH 9107
        msg = "Out of bounds nanosecond timestamp"
        with pytest.raises(OutOfBoundsDatetime, match=msg):
            to_datetime(dt_str, format="%Y%m%d")

    def test_to_datetime_utc(self):
        arr = np.array([parse("2012-06-13T01:39:00Z")], dtype=object)

        result = to_datetime(arr, utc=True)
        assert result.tz is timezone.utc

    def test_to_datetime_fixed_offset(self):
        from pandas.tests.indexes.datetimes.test_timezones import FixedOffset

        fixed_off = FixedOffset(-420, "-07:00")

        dates = [
            datetime(2000, 1, 1, tzinfo=fixed_off),
            datetime(2000, 1, 2, tzinfo=fixed_off),
            datetime(2000, 1, 3, tzinfo=fixed_off),
        ]
        result = to_datetime(dates)
        assert result.tz == fixed_off

    @pytest.mark.parametrize(
        "date",
        [
            ["2020-10-26 00:00:00+06:00", "2020-10-26 00:00:00+01:00"],
            ["2020-10-26 00:00:00+06:00", Timestamp("2018-01-01", tz="US/Pacific")],
            [
                "2020-10-26 00:00:00+06:00",
                datetime(2020, 1, 1, 18, tzinfo=pytz.timezone("Australia/Melbourne")),
            ],
        ],
    )
    def test_to_datetime_mixed_offsets_with_utc_false_deprecated(self, date):
        # GH 50887
        msg = "parsing datetimes with mixed time zones will raise an error"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            to_datetime(date, utc=False)


class TestToDatetimeUnit:
    @pytest.mark.parametrize("unit", ["Y", "M"])
    @pytest.mark.parametrize("item", [150, float(150)])
    def test_to_datetime_month_or_year_unit_int(self, cache, unit, item, request):
        # GH#50870 Note we have separate tests that pd.Timestamp gets these right
        ts = Timestamp(item, unit=unit)
        expected = DatetimeIndex([ts], dtype="M8[ns]")

        result = to_datetime([item], unit=unit, cache=cache)
        tm.assert_index_equal(result, expected)

        result = to_datetime(np.array([item], dtype=object), unit=unit, cache=cache)
        tm.assert_index_equal(result, expected)

        result = to_datetime(np.array([item]), unit=unit, cache=cache)
        tm.assert_index_equal(result, expected)

        # with a nan!
        result = to_datetime(np.array([item, np.nan]), unit=unit, cache=cache)
        assert result.isna()[1]
        tm.assert_index_equal(result[:1], expected)

    @pytest.mark.parametrize("unit", ["Y", "M"])
    def test_to_datetime_month_or_year_unit_non_round_float(self, cache, unit):
        # GH#50301
        # Match Timestamp behavior in disallowing non-round floats with
        #  Y or M unit
        warn_msg = "strings will be parsed as datetime strings"
        msg = f"Conversion of non-round float with unit={unit} is ambiguous"
        with pytest.raises(ValueError, match=msg):
            to_datetime([1.5], unit=unit, errors="raise")
        with pytest.raises(ValueError, match=msg):
            to_datetime(np.array([1.5]), unit=unit, errors="raise")
        with pytest.raises(ValueError, match=msg):
            with tm.assert_produces_warning(FutureWarning, match=warn_msg):
                to_datetime(["1.5"], unit=unit, errors="raise")

        # with errors="ignore" we also end up raising within the Timestamp
        #  constructor; this may not be ideal
        with pytest.raises(ValueError, match=msg):
            to_datetime([1.5], unit=unit, errors="ignore")

        res = to_datetime([1.5], unit=unit, errors="coerce")
        expected = Index([NaT], dtype="M8[ns]")
        tm.assert_index_equal(res, expected)

        with tm.assert_produces_warning(FutureWarning, match=warn_msg):
            res = to_datetime(["1.5"], unit=unit, errors="coerce")
        tm.assert_index_equal(res, expected)

        # round floats are OK
        res = to_datetime([1.0], unit=unit)
        expected = to_datetime([1], unit=unit)
        tm.assert_index_equal(res, expected)

    def test_unit(self, cache):
        # GH 11758
        # test proper behavior with errors
        msg = "cannot specify both format and unit"
        with pytest.raises(ValueError, match=msg):
            to_datetime([1], unit="D", format="%Y%m%d", cache=cache)

    def test_unit_str(self, cache):
        # GH 57051
        # Test that strs aren't dropping precision to 32-bit accidentally.
        with tm.assert_produces_warning(FutureWarning):
            res = to_datetime(["1704660000"], unit="s", origin="unix")
        expected = to_datetime([1704660000], unit="s", origin="unix")
        tm.assert_index_equal(res, expected)

    def test_unit_array_mixed_nans(self, cache):
        values = [11111111111111111, 1, 1.0, iNaT, NaT, np.nan, "NaT", ""]
        result = to_datetime(values, unit="D", errors="ignore", cache=cache)
        expected = Index(
            [
                11111111111111111,
                Timestamp("1970-01-02"),
                Timestamp("1970-01-02"),
                NaT,
                NaT,
                NaT,
                NaT,
                NaT,
            ],
            dtype=object,
        )
        tm.assert_index_equal(result, expected)

        result = to_datetime(values, unit="D", errors="coerce", cache=cache)
        expected = DatetimeIndex(
            ["NaT", "1970-01-02", "1970-01-02", "NaT", "NaT", "NaT", "NaT", "NaT"],
            dtype="M8[ns]",
        )
        tm.assert_index_equal(result, expected)

        msg = "cannot convert input 11111111111111111 with the unit 'D'"
        with pytest.raises(OutOfBoundsDatetime, match=msg):
            to_datetime(values, unit="D", errors="raise", cache=cache)

    def test_unit_array_mixed_nans_large_int(self, cache):
        values = [1420043460000000000000000, iNaT, NaT, np.nan, "NaT"]

        result = to_datetime(values, errors="ignore", unit="s", cache=cache)
        expected = Index([1420043460000000000000000, NaT, NaT, NaT, NaT], dtype=object)
        tm.assert_index_equal(result, expected)

        result = to_datetime(values, errors="coerce", unit="s", cache=cache)
        expected = DatetimeIndex(["NaT", "NaT", "NaT", "NaT", "NaT"], dtype="M8[ns]")
        tm.assert_index_equal(result, expected)

        msg = "cannot convert input 1420043460000000000000000 with the unit 's'"
        with pytest.raises(OutOfBoundsDatetime, match=msg):
            to_datetime(values, errors="raise", unit="s", cache=cache)

    def test_to_datetime_invalid_str_not_out_of_bounds_valuerror(self, cache):
        # if we have a string, then we raise a ValueError
        # and NOT an OutOfBoundsDatetime
        msg = "non convertible value foo with the unit 's'"
        with pytest.raises(ValueError, match=msg):
            to_datetime("foo", errors="raise", unit="s", cache=cache)

    @pytest.mark.parametrize("error", ["raise", "coerce", "ignore"])
    def test_unit_consistency(self, cache, error):
        # consistency of conversions
        expected = Timestamp("1970-05-09 14:25:11")
        result = to_datetime(11111111, unit="s", errors=error, cache=cache)
        assert result == expected
        assert isinstance(result, Timestamp)

    @pytest.mark.parametrize("errors", ["ignore", "raise", "coerce"])
    @pytest.mark.parametrize("dtype", ["float64", "int64"])
    def test_unit_with_numeric(self, cache, errors, dtype):
        # GH 13180
        # coercions from floats/ints are ok
        expected = DatetimeIndex(
            ["2015-06-19 05:33:20", "2015-05-27 22:33:20"], dtype="M8[ns]"
        )
        arr = np.array([1.434692e18, 1.432766e18]).astype(dtype)
        result = to_datetime(arr, errors=errors, cache=cache)
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "exp, arr, warning",
        [
            [
                ["NaT", "2015-06-19 05:33:20", "2015-05-27 22:33:20"],
                ["foo", 1.434692e18, 1.432766e18],
                UserWarning,
            ],
            [
                ["2015-06-19 05:33:20", "2015-05-27 22:33:20", "NaT", "NaT"],
                [1.434692e18, 1.432766e18, "foo", "NaT"],
                None,
            ],
        ],
    )
    def test_unit_with_numeric_coerce(self, cache, exp, arr, warning):
        # but we want to make sure that we are coercing
        # if we have ints/strings
        expected = DatetimeIndex(exp, dtype="M8[ns]")
        with tm.assert_produces_warning(warning, match="Could not infer format"):
            result = to_datetime(arr, errors="coerce", cache=cache)
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "arr",
        [
            [Timestamp("20130101"), 1.434692e18, 1.432766e18],
            [1.434692e18, 1.432766e18, Timestamp("20130101")],
        ],
    )
    def test_unit_mixed(self, cache, arr):
        # GH#50453 pre-2.0 with mixed numeric/datetimes and errors="coerce"
        #  the numeric entries would be coerced to NaT, was never clear exactly
        #  why.
        # mixed integers/datetimes
        expected = Index([Timestamp(x) for x in arr], dtype="M8[ns]")
        result = to_datetime(arr, errors="coerce", cache=cache)
        tm.assert_index_equal(result, expected)

        # GH#49037 pre-2.0 this raised, but it always worked with Series,
        #  was never clear why it was disallowed
        result = to_datetime(arr, errors="raise", cache=cache)
        tm.assert_index_equal(result, expected)

        result = DatetimeIndex(arr)
        tm.assert_index_equal(result, expected)

    def test_unit_rounding(self, cache):
        # GH 14156 & GH 20445: argument will incur floating point errors
        # but no premature rounding
        value = 1434743731.8770001
        result = to_datetime(value, unit="s", cache=cache)
        expected = Timestamp("2015-06-19 19:55:31.877000093")
        assert result == expected

        alt = Timestamp(value, unit="s")
        assert alt == result

    def test_unit_ignore_keeps_name(self, cache):
        # GH 21697
        expected = Index([15e9] * 2, name="name")
        result = to_datetime(expected, errors="ignore", unit="s", cache=cache)
        tm.assert_index_equal(result, expected)

    def test_to_datetime_errors_ignore_utc_true(self):
        # GH#23758
        result = to_datetime([1], unit="s", utc=True, errors="ignore")
        expected = DatetimeIndex(["1970-01-01 00:00:01"], dtype="M8[ns, UTC]")
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize("dtype", [int, float])
    def test_to_datetime_unit(self, dtype):
        epoch = 1370745748
        ser = Series([epoch + t for t in range(20)]).astype(dtype)
        result = to_datetime(ser, unit="s")
        expected = Series(
            [
                Timestamp("2013-06-09 02:42:28") + timedelta(seconds=t)
                for t in range(20)
            ],
            dtype="M8[ns]",
        )
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize("null", [iNaT, np.nan])
    def test_to_datetime_unit_with_nulls(self, null):
        epoch = 1370745748
        ser = Series([epoch + t for t in range(20)] + [null])
        result = to_datetime(ser, unit="s")
        expected = Series(
            [Timestamp("2013-06-09 02:42:28") + timedelta(seconds=t) for t in range(20)]
            + [NaT],
            dtype="M8[ns]",
        )
        tm.assert_series_equal(result, expected)

    def test_to_datetime_unit_fractional_seconds(self):
        # GH13834
        epoch = 1370745748
        ser = Series([epoch + t for t in np.arange(0, 2, 0.25)] + [iNaT]).astype(float)
        result = to_datetime(ser, unit="s")
        expected = Series(
            [
                Timestamp("2013-06-09 02:42:28") + timedelta(seconds=t)
                for t in np.arange(0, 2, 0.25)
            ]
            + [NaT],
            dtype="M8[ns]",
        )
        # GH20455 argument will incur floating point errors but no premature rounding
        result = result.round("ms")
        tm.assert_series_equal(result, expected)

    def test_to_datetime_unit_na_values(self):
        result = to_datetime([1, 2, "NaT", NaT, np.nan], unit="D")
        expected = DatetimeIndex(
            [Timestamp("1970-01-02"), Timestamp("1970-01-03")] + ["NaT"] * 3,
            dtype="M8[ns]",
        )
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize("bad_val", ["foo", 111111111])
    def test_to_datetime_unit_invalid(self, bad_val):
        msg = f"{bad_val} with the unit 'D'"
        with pytest.raises(ValueError, match=msg):
            to_datetime([1, 2, bad_val], unit="D")

    @pytest.mark.parametrize("bad_val", ["foo", 111111111])
    def test_to_timestamp_unit_coerce(self, bad_val):
        # coerce we can process
        expected = DatetimeIndex(
            [Timestamp("1970-01-02"), Timestamp("1970-01-03")] + ["NaT"] * 1,
            dtype="M8[ns]",
        )
        result = to_datetime([1, 2, bad_val], unit="D", errors="coerce")
        tm.assert_index_equal(result, expected)

    def test_float_to_datetime_raise_near_bounds(self):
        # GH50183
        msg = "cannot convert input with unit 'D'"
        oneday_in_ns = 1e9 * 60 * 60 * 24
        tsmax_in_days = 2**63 / oneday_in_ns  # 2**63 ns, in days
        # just in bounds
        should_succeed = Series(
            [0, tsmax_in_days - 0.005, -tsmax_in_days + 0.005], dtype=float
        )
        expected = (should_succeed * oneday_in_ns).astype(np.int64)
        for error_mode in ["raise", "coerce", "ignore"]:
            result1 = to_datetime(should_succeed, unit="D", errors=error_mode)
            # Cast to `np.float64` so that `rtol` and inexact checking kick in
            # (`check_exact` doesn't take place for integer dtypes)
            tm.assert_almost_equal(
                result1.astype(np.int64).astype(np.float64),
                expected.astype(np.float64),
                rtol=1e-10,
            )
        # just out of bounds
        should_fail1 = Series([0, tsmax_in_days + 0.005], dtype=float)
        should_fail2 = Series([0, -tsmax_in_days - 0.005], dtype=float)
        with pytest.raises(OutOfBoundsDatetime, match=msg):
            to_datetime(should_fail1, unit="D", errors="raise")
        with pytest.raises(OutOfBoundsDatetime, match=msg):
            to_datetime(should_fail2, unit="D", errors="raise")


class TestToDatetimeDataFrame:
    @pytest.fixture
    def df(self):
        return DataFrame(
            {
                "year": [2015, 2016],
                "month": [2, 3],
                "day": [4, 5],
                "hour": [6, 7],
                "minute": [58, 59],
                "second": [10, 11],
                "ms": [1, 1],
                "us": [2, 2],
                "ns": [3, 3],
            }
        )

    def test_dataframe(self, df, cache):
        result = to_datetime(
            {"year": df["year"], "month": df["month"], "day": df["day"]}, cache=cache
        )
        expected = Series(
            [Timestamp("20150204 00:00:00"), Timestamp("20160305 00:0:00")]
        )
        tm.assert_series_equal(result, expected)

        # dict-like
        result = to_datetime(df[["year", "month", "day"]].to_dict(), cache=cache)
        tm.assert_series_equal(result, expected)

    def test_dataframe_dict_with_constructable(self, df, cache):
        # dict but with constructable
        df2 = df[["year", "month", "day"]].to_dict()
        df2["month"] = 2
        result = to_datetime(df2, cache=cache)
        expected2 = Series(
            [Timestamp("20150204 00:00:00"), Timestamp("20160205 00:0:00")]
        )
        tm.assert_series_equal(result, expected2)

    @pytest.mark.parametrize(
        "unit",
        [
            {
                "year": "years",
                "month": "months",
                "day": "days",
                "hour": "hours",
                "minute": "minutes",
                "second": "seconds",
            },
            {
                "year": "year",
                "month": "month",
                "day": "day",
                "hour": "hour",
                "minute": "minute",
                "second": "second",
            },
        ],
    )
    def test_dataframe_field_aliases_column_subset(self, df, cache, unit):
        # unit mappings
        result = to_datetime(df[list(unit.keys())].rename(columns=unit), cache=cache)
        expected = Series(
            [Timestamp("20150204 06:58:10"), Timestamp("20160305 07:59:11")],
            dtype="M8[ns]",
        )
        tm.assert_series_equal(result, expected)

    def test_dataframe_field_aliases(self, df, cache):
        d = {
            "year": "year",
            "month": "month",
            "day": "day",
            "hour": "hour",
            "minute": "minute",
            "second": "second",
            "ms": "ms",
            "us": "us",
            "ns": "ns",
        }

        result = to_datetime(df.rename(columns=d), cache=cache)
        expected = Series(
            [
                Timestamp("20150204 06:58:10.001002003"),
                Timestamp("20160305 07:59:11.001002003"),
            ]
        )
        tm.assert_series_equal(result, expected)

    def test_dataframe_str_dtype(self, df, cache):
        # coerce back to int
        result = to_datetime(df.astype(str), cache=cache)
        expected = Series(
            [
                Timestamp("20150204 06:58:10.001002003"),
                Timestamp("20160305 07:59:11.001002003"),
            ]
        )
        tm.assert_series_equal(result, expected)

    def test_dataframe_coerce(self, cache):
        # passing coerce
        df2 = DataFrame({"year": [2015, 2016], "month": [2, 20], "day": [4, 5]})

        msg = (
            r'^cannot assemble the datetimes: time data ".+" doesn\'t '
            r'match format "%Y%m%d", at position 1\.'
        )
        with pytest.raises(ValueError, match=msg):
            to_datetime(df2, cache=cache)

        result = to_datetime(df2, errors="coerce", cache=cache)
        expected = Series([Timestamp("20150204 00:00:00"), NaT])
        tm.assert_series_equal(result, expected)

    def test_dataframe_extra_keys_raisesm(self, df, cache):
        # extra columns
        msg = r"extra keys have been passed to the datetime assemblage: \[foo\]"
        with pytest.raises(ValueError, match=msg):
            df2 = df.copy()
            df2["foo"] = 1
            to_datetime(df2, cache=cache)

    @pytest.mark.parametrize(
        "cols",
        [
            ["year"],
            ["year", "month"],
            ["year", "month", "second"],
            ["month", "day"],
            ["year", "day", "second"],
        ],
    )
    def test_dataframe_missing_keys_raises(self, df, cache, cols):
        # not enough
        msg = (
            r"to assemble mappings requires at least that \[year, month, "
            r"day\] be specified: \[.+\] is missing"
        )
        with pytest.raises(ValueError, match=msg):
            to_datetime(df[cols], cache=cache)

    def test_dataframe_duplicate_columns_raises(self, cache):
        # duplicates
        msg = "cannot assemble with duplicate keys"
        df2 = DataFrame({"year": [2015, 2016], "month": [2, 20], "day": [4, 5]})
        df2.columns = ["year", "year", "day"]
        with pytest.raises(ValueError, match=msg):
            to_datetime(df2, cache=cache)

        df2 = DataFrame(
            {"year": [2015, 2016], "month": [2, 20], "day": [4, 5], "hour": [4, 5]}
        )
        df2.columns = ["year", "month", "day", "day"]
        with pytest.raises(ValueError, match=msg):
            to_datetime(df2, cache=cache)

    def test_dataframe_int16(self, cache):
        # GH#13451
        df = DataFrame({"year": [2015, 2016], "month": [2, 3], "day": [4, 5]})

        # int16
        result = to_datetime(df.astype("int16"), cache=cache)
        expected = Series(
            [Timestamp("20150204 00:00:00"), Timestamp("20160305 00:00:00")]
        )
        tm.assert_series_equal(result, expected)

    def test_dataframe_mixed(self, cache):
        # mixed dtypes
        df = DataFrame({"year": [2015, 2016], "month": [2, 3], "day": [4, 5]})
        df["month"] = df["month"].astype("int8")
        df["day"] = df["day"].astype("int8")
        result = to_datetime(df, cache=cache)
        expected = Series(
            [Timestamp("20150204 00:00:00"), Timestamp("20160305 00:00:00")]
        )
        tm.assert_series_equal(result, expected)

    def test_dataframe_float(self, cache):
        # float
        df = DataFrame({"year": [2000, 2001], "month": [1.5, 1], "day": [1, 1]})
        msg = (
            r"^cannot assemble the datetimes: unconverted data remains when parsing "
            r'with format ".*": "1", at position 0.'
        )
        with pytest.raises(ValueError, match=msg):
            to_datetime(df, cache=cache)

    def test_dataframe_utc_true(self):
        # GH#23760
        df = DataFrame({"year": [2015, 2016], "month": [2, 3], "day": [4, 5]})
        result = to_datetime(df, utc=True)
        expected = Series(
            np.array(["2015-02-04", "2016-03-05"], dtype="datetime64[ns]")
        ).dt.tz_localize("UTC")
        tm.assert_series_equal(result, expected)


class TestToDatetimeMisc:
    def test_to_datetime_barely_out_of_bounds(self):
        # GH#19529
        # GH#19382 close enough to bounds that dropping nanos would result
        # in an in-bounds datetime
        arr = np.array(["2262-04-11 23:47:16.854775808"], dtype=object)

        msg = "^Out of bounds nanosecond timestamp: .*, at position 0"
        with pytest.raises(OutOfBoundsDatetime, match=msg):
            to_datetime(arr)

    @pytest.mark.parametrize(
        "arg, exp_str",
        [
            ["2012-01-01 00:00:00", "2012-01-01 00:00:00"],
            ["20121001", "2012-10-01"],  # bad iso 8601
        ],
    )
    def test_to_datetime_iso8601(self, cache, arg, exp_str):
        result = to_datetime([arg], cache=cache)
        exp = Timestamp(exp_str)
        assert result[0] == exp

    @pytest.mark.parametrize(
        "input, format",
        [
            ("2012", "%Y-%m"),
            ("2012-01", "%Y-%m-%d"),
            ("2012-01-01", "%Y-%m-%d %H"),
            ("2012-01-01 10", "%Y-%m-%d %H:%M"),
            ("2012-01-01 10:00", "%Y-%m-%d %H:%M:%S"),
            ("2012-01-01 10:00:00", "%Y-%m-%d %H:%M:%S.%f"),
            ("2012-01-01 10:00:00.123", "%Y-%m-%d %H:%M:%S.%f%z"),
            (0, "%Y-%m-%d"),
        ],
    )
    @pytest.mark.parametrize("exact", [True, False])
    def test_to_datetime_iso8601_fails(self, input, format, exact):
        # https://github.com/pandas-dev/pandas/issues/12649
        # `format` is longer than the string, so this fails regardless of `exact`
        with pytest.raises(
            ValueError,
            match=(
                rf"time data \"{input}\" doesn't match format "
                rf"\"{format}\", at position 0"
            ),
        ):
            to_datetime(input, format=format, exact=exact)

    @pytest.mark.parametrize(
        "input, format",
        [
            ("2012-01-01", "%Y-%m"),
            ("2012-01-01 10", "%Y-%m-%d"),
            ("2012-01-01 10:00", "%Y-%m-%d %H"),
            ("2012-01-01 10:00:00", "%Y-%m-%d %H:%M"),
            (0, "%Y-%m-%d"),
        ],
    )
    def test_to_datetime_iso8601_exact_fails(self, input, format):
        # https://github.com/pandas-dev/pandas/issues/12649
        # `format` is shorter than the date string, so only fails with `exact=True`
        msg = "|".join(
            [
                '^unconverted data remains when parsing with format ".*": ".*"'
                f", at position 0. {PARSING_ERR_MSG}$",
                f'^time data ".*" doesn\'t match format ".*", at position 0. '
                f"{PARSING_ERR_MSG}$",
            ]
        )
        with pytest.raises(
            ValueError,
            match=(msg),
        ):
            to_datetime(input, format=format)

    @pytest.mark.parametrize(
        "input, format",
        [
            ("2012-01-01", "%Y-%m"),
            ("2012-01-01 00", "%Y-%m-%d"),
            ("2012-01-01 00:00", "%Y-%m-%d %H"),
            ("2012-01-01 00:00:00", "%Y-%m-%d %H:%M"),
        ],
    )
    def test_to_datetime_iso8601_non_exact(self, input, format):
        # https://github.com/pandas-dev/pandas/issues/12649
        expected = Timestamp(2012, 1, 1)
        result = to_datetime(input, format=format, exact=False)
        assert result == expected

    @pytest.mark.parametrize(
        "input, format",
        [
            ("2020-01", "%Y/%m"),
            ("2020-01-01", "%Y/%m/%d"),
            ("2020-01-01 00", "%Y/%m/%dT%H"),
            ("2020-01-01T00", "%Y/%m/%d %H"),
            ("2020-01-01 00:00", "%Y/%m/%dT%H:%M"),
            ("2020-01-01T00:00", "%Y/%m/%d %H:%M"),
            ("2020-01-01 00:00:00", "%Y/%m/%dT%H:%M:%S"),
            ("2020-01-01T00:00:00", "%Y/%m/%d %H:%M:%S"),
        ],
    )
    def test_to_datetime_iso8601_separator(self, input, format):
        # https://github.com/pandas-dev/pandas/issues/12649
        with pytest.raises(
            ValueError,
            match=(
                rf"time data \"{input}\" doesn\'t match format "
                rf"\"{format}\", at position 0"
            ),
        ):
            to_datetime(input, format=format)

    @pytest.mark.parametrize(
        "input, format",
        [
            ("2020-01", "%Y-%m"),
            ("2020-01-01", "%Y-%m-%d"),
            ("2020-01-01 00", "%Y-%m-%d %H"),
            ("2020-01-01T00", "%Y-%m-%dT%H"),
            ("2020-01-01 00:00", "%Y-%m-%d %H:%M"),
            ("2020-01-01T00:00", "%Y-%m-%dT%H:%M"),
            ("2020-01-01 00:00:00", "%Y-%m-%d %H:%M:%S"),
            ("2020-01-01T00:00:00", "%Y-%m-%dT%H:%M:%S"),
            ("2020-01-01T00:00:00.000", "%Y-%m-%dT%H:%M:%S.%f"),
            ("2020-01-01T00:00:00.000000", "%Y-%m-%dT%H:%M:%S.%f"),
            ("2020-01-01T00:00:00.000000000", "%Y-%m-%dT%H:%M:%S.%f"),
        ],
    )
    def test_to_datetime_iso8601_valid(self, input, format):
        # https://github.com/pandas-dev/pandas/issues/12649
        expected = Timestamp(2020, 1, 1)
        result = to_datetime(input, format=format)
        assert result == expected

    @pytest.mark.parametrize(
        "input, format",
        [
            ("2020-1", "%Y-%m"),
            ("2020-1-1", "%Y-%m-%d"),
            ("2020-1-1 0", "%Y-%m-%d %H"),
            ("2020-1-1T0", "%Y-%m-%dT%H"),
            ("2020-1-1 0:0", "%Y-%m-%d %H:%M"),
            ("2020-1-1T0:0", "%Y-%m-%dT%H:%M"),
            ("2020-1-1 0:0:0", "%Y-%m-%d %H:%M:%S"),
            ("2020-1-1T0:0:0", "%Y-%m-%dT%H:%M:%S"),
            ("2020-1-1T0:0:0.000", "%Y-%m-%dT%H:%M:%S.%f"),
            ("2020-1-1T0:0:0.000000", "%Y-%m-%dT%H:%M:%S.%f"),
            ("2020-1-1T0:0:0.000000000", "%Y-%m-%dT%H:%M:%S.%f"),
        ],
    )
    def test_to_datetime_iso8601_non_padded(self, input, format):
        # https://github.com/pandas-dev/pandas/issues/21422
        expected = Timestamp(2020, 1, 1)
        result = to_datetime(input, format=format)
        assert result == expected

    @pytest.mark.parametrize(
        "input, format",
        [
            ("2020-01-01T00:00:00.000000000+00:00", "%Y-%m-%dT%H:%M:%S.%f%z"),
            ("2020-01-01T00:00:00+00:00", "%Y-%m-%dT%H:%M:%S%z"),
            ("2020-01-01T00:00:00Z", "%Y-%m-%dT%H:%M:%S%z"),
        ],
    )
    def test_to_datetime_iso8601_with_timezone_valid(self, input, format):
        # https://github.com/pandas-dev/pandas/issues/12649
        expected = Timestamp(2020, 1, 1, tzinfo=pytz.UTC)
        result = to_datetime(input, format=format)
        assert result == expected

    def test_to_datetime_default(self, cache):
        rs = to_datetime("2001", cache=cache)
        xp = datetime(2001, 1, 1)
        assert rs == xp

    @pytest.mark.xfail(reason="fails to enforce dayfirst=True, which would raise")
    def test_to_datetime_respects_dayfirst(self, cache):
        # dayfirst is essentially broken

        # The msg here is not important since it isn't actually raised yet.
        msg = "Invalid date specified"
        with pytest.raises(ValueError, match=msg):
            # if dayfirst is respected, then this would parse as month=13, which
            #  would raise
            with tm.assert_produces_warning(UserWarning, match="Provide format"):
                to_datetime("01-13-2012", dayfirst=True, cache=cache)

    def test_to_datetime_on_datetime64_series(self, cache):
        # #2699
        ser = Series(date_range("1/1/2000", periods=10))

        result = to_datetime(ser, cache=cache)
        assert result[0] == ser[0]

    def test_to_datetime_with_space_in_series(self, cache):
        # GH 6428
        ser = Series(["10/18/2006", "10/18/2008", " "])
        msg = (
            r'^time data " " doesn\'t match format "%m/%d/%Y", '
            rf"at position 2. {PARSING_ERR_MSG}$"
        )
        with pytest.raises(ValueError, match=msg):
            to_datetime(ser, errors="raise", cache=cache)
        result_coerce = to_datetime(ser, errors="coerce", cache=cache)
        expected_coerce = Series([datetime(2006, 10, 18), datetime(2008, 10, 18), NaT])
        tm.assert_series_equal(result_coerce, expected_coerce)
        result_ignore = to_datetime(ser, errors="ignore", cache=cache)
        tm.assert_series_equal(result_ignore, ser)

    @td.skip_if_not_us_locale
    def test_to_datetime_with_apply(self, cache):
        # this is only locale tested with US/None locales
        # GH 5195
        # with a format and coerce a single item to_datetime fails
        td = Series(["May 04", "Jun 02", "Dec 11"], index=[1, 2, 3])
        expected = to_datetime(td, format="%b %y", cache=cache)
        result = td.apply(to_datetime, format="%b %y", cache=cache)
        tm.assert_series_equal(result, expected)

    def test_to_datetime_timezone_name(self):
        # https://github.com/pandas-dev/pandas/issues/49748
        result = to_datetime("2020-01-01 00:00:00UTC", format="%Y-%m-%d %H:%M:%S%Z")
        expected = Timestamp(2020, 1, 1).tz_localize("UTC")
        assert result == expected

    @td.skip_if_not_us_locale
    @pytest.mark.parametrize("errors", ["raise", "coerce", "ignore"])
    def test_to_datetime_with_apply_with_empty_str(self, cache, errors):
        # this is only locale tested with US/None locales
        # GH 5195, GH50251
        # with a format and coerce a single item to_datetime fails
        td = Series(["May 04", "Jun 02", ""], index=[1, 2, 3])
        expected = to_datetime(td, format="%b %y", errors=errors, cache=cache)

        result = td.apply(
            lambda x: to_datetime(x, format="%b %y", errors="coerce", cache=cache)
        )
        tm.assert_series_equal(result, expected)

    def test_to_datetime_empty_stt(self, cache):
        # empty string
        result = to_datetime("", cache=cache)
        assert result is NaT

    def test_to_datetime_empty_str_list(self, cache):
        result = to_datetime(["", ""], cache=cache)
        assert isna(result).all()

    def test_to_datetime_zero(self, cache):
        # ints
        result = Timestamp(0)
        expected = to_datetime(0, cache=cache)
        assert result == expected

    def test_to_datetime_strings(self, cache):
        # GH 3888 (strings)
        expected = to_datetime(["2012"], cache=cache)[0]
        result = to_datetime("2012", cache=cache)
        assert result == expected

    def test_to_datetime_strings_variation(self, cache):
        array = ["2012", "20120101", "20120101 12:01:01"]
        expected = [to_datetime(dt_str, cache=cache) for dt_str in array]
        result = [Timestamp(date_str) for date_str in array]
        tm.assert_almost_equal(result, expected)

    @pytest.mark.parametrize("result", [Timestamp("2012"), to_datetime("2012")])
    def test_to_datetime_strings_vs_constructor(self, result):
        expected = Timestamp(2012, 1, 1)
        assert result == expected

    def test_to_datetime_unprocessable_input(self, cache):
        # GH 4928
        # GH 21864
        result = to_datetime([1, "1"], errors="ignore", cache=cache)

        expected = Index(np.array([1, "1"], dtype="O"))
        tm.assert_equal(result, expected)
        msg = '^Given date string "1" not likely a datetime, at position 1$'
        with pytest.raises(ValueError, match=msg):
            to_datetime([1, "1"], errors="raise", cache=cache)

    def test_to_datetime_unhashable_input(self, cache):
        series = Series([["a"]] * 100)
        result = to_datetime(series, errors="ignore", cache=cache)
        tm.assert_series_equal(series, result)

    def test_to_datetime_other_datetime64_units(self):
        # 5/25/2012
        scalar = np.int64(1337904000000000).view("M8[us]")
        as_obj = scalar.astype("O")

        index = DatetimeIndex([scalar])
        assert index[0] == scalar.astype("O")

        value = Timestamp(scalar)
        assert value == as_obj

    def test_to_datetime_list_of_integers(self):
        rng = date_range("1/1/2000", periods=20)
        rng = DatetimeIndex(rng.values)

        ints = list(rng.asi8)

        result = DatetimeIndex(ints)

        tm.assert_index_equal(rng, result)

    def test_to_datetime_overflow(self):
        # gh-17637
        # we are overflowing Timedelta range here
        msg = "Cannot cast 139999 days 00:00:00 to unit='ns' without overflow"
        with pytest.raises(OutOfBoundsTimedelta, match=msg):
            date_range(start="1/1/1700", freq="B", periods=100000)

    def test_string_invalid_operation(self, cache):
        invalid = np.array(["87156549591102612381000001219H5"], dtype=object)
        # GH #51084

        with pytest.raises(ValueError, match="Unknown datetime string format"):
            to_datetime(invalid, errors="raise", cache=cache)

    def test_string_na_nat_conversion(self, cache):
        # GH #999, #858

        strings = np.array(["1/1/2000", "1/2/2000", np.nan, "1/4/2000"], dtype=object)

        expected = np.empty(4, dtype="M8[ns]")
        for i, val in enumerate(strings):
            if isna(val):
                expected[i] = iNaT
            else:
                expected[i] = parse(val)

        result = tslib.array_to_datetime(strings)[0]
        tm.assert_almost_equal(result, expected)

        result2 = to_datetime(strings, cache=cache)
        assert isinstance(result2, DatetimeIndex)
        tm.assert_numpy_array_equal(result, result2.values)

    def test_string_na_nat_conversion_malformed(self, cache):
        malformed = np.array(["1/100/2000", np.nan], dtype=object)

        # GH 10636, default is now 'raise'
        msg = r"Unknown datetime string format"
        with pytest.raises(ValueError, match=msg):
            to_datetime(malformed, errors="raise", cache=cache)

        result = to_datetime(malformed, errors="ignore", cache=cache)
        # GH 21864
        expected = Index(malformed, dtype=object)
        tm.assert_index_equal(result, expected)

        with pytest.raises(ValueError, match=msg):
            to_datetime(malformed, errors="raise", cache=cache)

    def test_string_na_nat_conversion_with_name(self, cache):
        idx = ["a", "b", "c", "d", "e"]
        series = Series(
            ["1/1/2000", np.nan, "1/3/2000", np.nan, "1/5/2000"], index=idx, name="foo"
        )
        dseries = Series(
            [
                to_datetime("1/1/2000", cache=cache),
                np.nan,
                to_datetime("1/3/2000", cache=cache),
                np.nan,
                to_datetime("1/5/2000", cache=cache),
            ],
            index=idx,
            name="foo",
        )

        result = to_datetime(series, cache=cache)
        dresult = to_datetime(dseries, cache=cache)

        expected = Series(np.empty(5, dtype="M8[ns]"), index=idx)
        for i in range(5):
            x = series.iloc[i]
            if isna(x):
                expected.iloc[i] = NaT
            else:
                expected.iloc[i] = to_datetime(x, cache=cache)

        tm.assert_series_equal(result, expected, check_names=False)
        assert result.name == "foo"

        tm.assert_series_equal(dresult, expected, check_names=False)
        assert dresult.name == "foo"

    @pytest.mark.parametrize(
        "unit",
        ["h", "m", "s", "ms", "us", "ns"],
    )
    def test_dti_constructor_numpy_timeunits(self, cache, unit):
        # GH 9114
        dtype = np.dtype(f"M8[{unit}]")
        base = to_datetime(["2000-01-01T00:00", "2000-01-02T00:00", "NaT"], cache=cache)

        values = base.values.astype(dtype)

        if unit in ["h", "m"]:
            # we cast to closest supported unit
            unit = "s"
        exp_dtype = np.dtype(f"M8[{unit}]")
        expected = DatetimeIndex(base.astype(exp_dtype))
        assert expected.dtype == exp_dtype

        tm.assert_index_equal(DatetimeIndex(values), expected)
        tm.assert_index_equal(to_datetime(values, cache=cache), expected)

    def test_dayfirst(self, cache):
        # GH 5917
        arr = ["10/02/2014", "11/02/2014", "12/02/2014"]
        expected = DatetimeIndex(
            [datetime(2014, 2, 10), datetime(2014, 2, 11), datetime(2014, 2, 12)]
        )
        idx1 = DatetimeIndex(arr, dayfirst=True)
        idx2 = DatetimeIndex(np.array(arr), dayfirst=True)
        idx3 = to_datetime(arr, dayfirst=True, cache=cache)
        idx4 = to_datetime(np.array(arr), dayfirst=True, cache=cache)
        idx5 = DatetimeIndex(Index(arr), dayfirst=True)
        idx6 = DatetimeIndex(Series(arr), dayfirst=True)
        tm.assert_index_equal(expected, idx1)
        tm.assert_index_equal(expected, idx2)
        tm.assert_index_equal(expected, idx3)
        tm.assert_index_equal(expected, idx4)
        tm.assert_index_equal(expected, idx5)
        tm.assert_index_equal(expected, idx6)

    def test_dayfirst_warnings_valid_input(self):
        # GH 12585
        warning_msg = (
            "Parsing dates in .* format when dayfirst=.* was specified. "
            "Pass `dayfirst=.*` or specify a format to silence this warning."
        )

        # CASE 1: valid input
        arr = ["31/12/2014", "10/03/2011"]
        expected = DatetimeIndex(
            ["2014-12-31", "2011-03-10"], dtype="datetime64[ns]", freq=None
        )

        # A. dayfirst arg correct, no warning
        res1 = to_datetime(arr, dayfirst=True)
        tm.assert_index_equal(expected, res1)

        # B. dayfirst arg incorrect, warning
        with tm.assert_produces_warning(UserWarning, match=warning_msg):
            res2 = to_datetime(arr, dayfirst=False)
        tm.assert_index_equal(expected, res2)

    def test_dayfirst_warnings_invalid_input(self):
        # CASE 2: invalid input
        # cannot consistently process with single format
        # ValueError *always* raised

        # first in DD/MM/YYYY, second in MM/DD/YYYY
        arr = ["31/12/2014", "03/30/2011"]

        with pytest.raises(
            ValueError,
            match=(
                r'^time data "03/30/2011" doesn\'t match format '
                rf'"%d/%m/%Y", at position 1. {PARSING_ERR_MSG}$'
            ),
        ):
            to_datetime(arr, dayfirst=True)

    @pytest.mark.parametrize("klass", [DatetimeIndex, DatetimeArray._from_sequence])
    def test_to_datetime_dta_tz(self, klass):
        # GH#27733
        dti = date_range("2015-04-05", periods=3).rename("foo")
        expected = dti.tz_localize("UTC")

        obj = klass(dti)
        expected = klass(expected)

        result = to_datetime(obj, utc=True)
        tm.assert_equal(result, expected)


class TestGuessDatetimeFormat:
    @pytest.mark.parametrize(
        "test_list",
        [
            [
                "2011-12-30 00:00:00.000000",
                "2011-12-30 00:00:00.000000",
                "2011-12-30 00:00:00.000000",
            ],
            [np.nan, np.nan, "2011-12-30 00:00:00.000000"],
            ["", "2011-12-30 00:00:00.000000"],
            ["NaT", "2011-12-30 00:00:00.000000"],
            ["2011-12-30 00:00:00.000000", "random_string"],
            ["now", "2011-12-30 00:00:00.000000"],
            ["today", "2011-12-30 00:00:00.000000"],
        ],
    )
    def test_guess_datetime_format_for_array(self, test_list):
        expected_format = "%Y-%m-%d %H:%M:%S.%f"
        test_array = np.array(test_list, dtype=object)
        assert tools._guess_datetime_format_for_array(test_array) == expected_format

    @td.skip_if_not_us_locale
    def test_guess_datetime_format_for_array_all_nans(self):
        format_for_string_of_nans = tools._guess_datetime_format_for_array(
            np.array([np.nan, np.nan, np.nan], dtype="O")
        )
        assert format_for_string_of_nans is None


class TestToDatetimeInferFormat:
    @pytest.mark.parametrize(
        "test_format", ["%m-%d-%Y", "%m/%d/%Y %H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S.%f"]
    )
    def test_to_datetime_infer_datetime_format_consistent_format(
        self, cache, test_format
    ):
        ser = Series(date_range("20000101", periods=50, freq="h"))

        s_as_dt_strings = ser.apply(lambda x: x.strftime(test_format))

        with_format = to_datetime(s_as_dt_strings, format=test_format, cache=cache)
        without_format = to_datetime(s_as_dt_strings, cache=cache)

        # Whether the format is explicitly passed, or
        # it is inferred, the results should all be the same
        tm.assert_series_equal(with_format, without_format)

    def test_to_datetime_inconsistent_format(self, cache):
        data = ["01/01/2011 00:00:00", "01-02-2011 00:00:00", "2011-01-03T00:00:00"]
        ser = Series(np.array(data))
        msg = (
            r'^time data "01-02-2011 00:00:00" doesn\'t match format '
            rf'"%m/%d/%Y %H:%M:%S", at position 1. {PARSING_ERR_MSG}$'
        )
        with pytest.raises(ValueError, match=msg):
            to_datetime(ser, cache=cache)

    def test_to_datetime_consistent_format(self, cache):
        data = ["Jan/01/2011", "Feb/01/2011", "Mar/01/2011"]
        ser = Series(np.array(data))
        result = to_datetime(ser, cache=cache)
        expected = Series(
            ["2011-01-01", "2011-02-01", "2011-03-01"], dtype="datetime64[ns]"
        )
        tm.assert_series_equal(result, expected)

    def test_to_datetime_series_with_nans(self, cache):
        ser = Series(
            np.array(
                ["01/01/2011 00:00:00", np.nan, "01/03/2011 00:00:00", np.nan],
                dtype=object,
            )
        )
        result = to_datetime(ser, cache=cache)
        expected = Series(
            ["2011-01-01", NaT, "2011-01-03", NaT], dtype="datetime64[ns]"
        )
        tm.assert_series_equal(result, expected)

    def test_to_datetime_series_start_with_nans(self, cache):
        ser = Series(
            np.array(
                [
                    np.nan,
                    np.nan,
                    "01/01/2011 00:00:00",
                    "01/02/2011 00:00:00",
                    "01/03/2011 00:00:00",
                ],
                dtype=object,
            )
        )

        result = to_datetime(ser, cache=cache)
        expected = Series(
            [NaT, NaT, "2011-01-01", "2011-01-02", "2011-01-03"], dtype="datetime64[ns]"
        )
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize(
        "tz_name, offset",
        [("UTC", 0), ("UTC-3", 180), ("UTC+3", -180)],
    )
    def test_infer_datetime_format_tz_name(self, tz_name, offset):
        # GH 33133
        ser = Series([f"2019-02-02 08:07:13 {tz_name}"])
        result = to_datetime(ser)
        tz = timezone(timedelta(minutes=offset))
        expected = Series([Timestamp("2019-02-02 08:07:13").tz_localize(tz)])
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize(
        "ts,zero_tz",
        [
            ("2019-02-02 08:07:13", "Z"),
            ("2019-02-02 08:07:13", ""),
            ("2019-02-02 08:07:13.012345", "Z"),
            ("2019-02-02 08:07:13.012345", ""),
        ],
    )
    def test_infer_datetime_format_zero_tz(self, ts, zero_tz):
        # GH 41047
        ser = Series([ts + zero_tz])
        result = to_datetime(ser)
        tz = pytz.utc if zero_tz == "Z" else None
        expected = Series([Timestamp(ts, tz=tz)])
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize("format", [None, "%Y-%m-%d"])
    def test_to_datetime_iso8601_noleading_0s(self, cache, format):
        # GH 11871
        ser = Series(["2014-1-1", "2014-2-2", "2015-3-3"])
        expected = Series(
            [
                Timestamp("2014-01-01"),
                Timestamp("2014-02-02"),
                Timestamp("2015-03-03"),
            ]
        )
        result = to_datetime(ser, format=format, cache=cache)
        tm.assert_series_equal(result, expected)

    def test_parse_dates_infer_datetime_format_warning(self):
        # GH 49024
        with tm.assert_produces_warning(
            UserWarning,
            match="The argument 'infer_datetime_format' is deprecated",
        ):
            to_datetime(["10-10-2000"], infer_datetime_format=True)


class TestDaysInMonth:
    # tests for issue #10154

    @pytest.mark.parametrize(
        "arg, format",
        [
            ["2015-02-29", None],
            ["2015-02-29", "%Y-%m-%d"],
            ["2015-02-32", "%Y-%m-%d"],
            ["2015-04-31", "%Y-%m-%d"],
        ],
    )
    def test_day_not_in_month_coerce(self, cache, arg, format):
        assert isna(to_datetime(arg, errors="coerce", format=format, cache=cache))

    def test_day_not_in_month_raise(self, cache):
        msg = "day is out of range for month: 2015-02-29, at position 0"
        with pytest.raises(ValueError, match=msg):
            to_datetime("2015-02-29", errors="raise", cache=cache)

    @pytest.mark.parametrize(
        "arg, format, msg",
        [
            (
                "2015-02-29",
                "%Y-%m-%d",
                f"^day is out of range for month, at position 0. {PARSING_ERR_MSG}$",
            ),
            (
                "2015-29-02",
                "%Y-%d-%m",
                f"^day is out of range for month, at position 0. {PARSING_ERR_MSG}$",
            ),
            (
                "2015-02-32",
                "%Y-%m-%d",
                '^unconverted data remains when parsing with format "%Y-%m-%d": "2", '
                f"at position 0. {PARSING_ERR_MSG}$",
            ),
            (
                "2015-32-02",
                "%Y-%d-%m",
                '^time data "2015-32-02" doesn\'t match format "%Y-%d-%m", '
                f"at position 0. {PARSING_ERR_MSG}$",
            ),
            (
                "2015-04-31",
                "%Y-%m-%d",
                f"^day is out of range for month, at position 0. {PARSING_ERR_MSG}$",
            ),
            (
                "2015-31-04",
                "%Y-%d-%m",
                f"^day is out of range for month, at position 0. {PARSING_ERR_MSG}$",
            ),
        ],
    )
    def test_day_not_in_month_raise_value(self, cache, arg, format, msg):
        # https://github.com/pandas-dev/pandas/issues/50462
        with pytest.raises(ValueError, match=msg):
            to_datetime(arg, errors="raise", format=format, cache=cache)

    @pytest.mark.parametrize(
        "expected, format",
        [
            ["2015-02-29", None],
            ["2015-02-29", "%Y-%m-%d"],
            ["2015-02-29", "%Y-%m-%d"],
            ["2015-04-31", "%Y-%m-%d"],
        ],
    )
    def test_day_not_in_month_ignore(self, cache, expected, format):
        result = to_datetime(expected, errors="ignore", format=format, cache=cache)
        assert result == expected


class TestDatetimeParsingWrappers:
    @pytest.mark.parametrize(
        "date_str, expected",
        [
            ("2011-01-01", datetime(2011, 1, 1)),
            ("2Q2005", datetime(2005, 4, 1)),
            ("2Q05", datetime(2005, 4, 1)),
            ("2005Q1", datetime(2005, 1, 1)),
            ("05Q1", datetime(2005, 1, 1)),
            ("2011Q3", datetime(2011, 7, 1)),
            ("11Q3", datetime(2011, 7, 1)),
            ("3Q2011", datetime(2011, 7, 1)),
            ("3Q11", datetime(2011, 7, 1)),
            # quarterly without space
            ("2000Q4", datetime(2000, 10, 1)),
            ("00Q4", datetime(2000, 10, 1)),
            ("4Q2000", datetime(2000, 10, 1)),
            ("4Q00", datetime(2000, 10, 1)),
            ("2000q4", datetime(2000, 10, 1)),
            ("2000-Q4", datetime(2000, 10, 1)),
            ("00-Q4", datetime(2000, 10, 1)),
            ("4Q-2000", datetime(2000, 10, 1)),
            ("4Q-00", datetime(2000, 10, 1)),
            ("00q4", datetime(2000, 10, 1)),
            ("2005", datetime(2005, 1, 1)),
            ("2005-11", datetime(2005, 11, 1)),
            ("2005 11", datetime(2005, 11, 1)),
            ("11-2005", datetime(2005, 11, 1)),
            ("11 2005", datetime(2005, 11, 1)),
            ("200511", datetime(2020, 5, 11)),
            ("20051109", datetime(2005, 11, 9)),
            ("20051109 10:15", datetime(2005, 11, 9, 10, 15)),
            ("20051109 08H", datetime(2005, 11, 9, 8, 0)),
            ("2005-11-09 10:15", datetime(2005, 11, 9, 10, 15)),
            ("2005-11-09 08H", datetime(2005, 11, 9, 8, 0)),
            ("2005/11/09 10:15", datetime(2005, 11, 9, 10, 15)),
            ("2005/11/09 10:15:32", datetime(2005, 11, 9, 10, 15, 32)),
            ("2005/11/09 10:15:32 AM", datetime(2005, 11, 9, 10, 15, 32)),
            ("2005/11/09 10:15:32 PM", datetime(2005, 11, 9, 22, 15, 32)),
            ("2005/11/09 08H", datetime(2005, 11, 9, 8, 0)),
            ("Thu Sep 25 10:36:28 2003", datetime(2003, 9, 25, 10, 36, 28)),
            ("Thu Sep 25 2003", datetime(2003, 9, 25)),
            ("Sep 25 2003", datetime(2003, 9, 25)),
            ("January 1 2014", datetime(2014, 1, 1)),
            # GH#10537
            ("2014-06", datetime(2014, 6, 1)),
            ("06-2014", datetime(2014, 6, 1)),
            ("2014-6", datetime(2014, 6, 1)),
            ("6-2014", datetime(2014, 6, 1)),
            ("20010101 12", datetime(2001, 1, 1, 12)),
            ("20010101 1234", datetime(2001, 1, 1, 12, 34)),
            ("20010101 123456", datetime(2001, 1, 1, 12, 34, 56)),
        ],
    )
    def test_parsers(self, date_str, expected, cache):
        # dateutil >= 2.5.0 defaults to yearfirst=True
        # https://github.com/dateutil/dateutil/issues/217
        yearfirst = True

        result1, _ = parsing.parse_datetime_string_with_reso(
            date_str, yearfirst=yearfirst
        )
        result2 = to_datetime(date_str, yearfirst=yearfirst)
        result3 = to_datetime([date_str], yearfirst=yearfirst)
        # result5 is used below
        result4 = to_datetime(
            np.array([date_str], dtype=object), yearfirst=yearfirst, cache=cache
        )
        result6 = DatetimeIndex([date_str], yearfirst=yearfirst)
        # result7 is used below
        result8 = DatetimeIndex(Index([date_str]), yearfirst=yearfirst)
        result9 = DatetimeIndex(Series([date_str]), yearfirst=yearfirst)

        for res in [result1, result2]:
            assert res == expected
        for res in [result3, result4, result6, result8, result9]:
            exp = DatetimeIndex([Timestamp(expected)])
            tm.assert_index_equal(res, exp)

        # these really need to have yearfirst, but we don't support
        if not yearfirst:
            result5 = Timestamp(date_str)
            assert result5 == expected
            result7 = date_range(date_str, freq="S", periods=1, yearfirst=yearfirst)
            assert result7 == expected

    def test_na_values_with_cache(
        self, cache, unique_nulls_fixture, unique_nulls_fixture2
    ):
        # GH22305
        expected = Index([NaT, NaT], dtype="datetime64[ns]")
        result = to_datetime([unique_nulls_fixture, unique_nulls_fixture2], cache=cache)
        tm.assert_index_equal(result, expected)

    def test_parsers_nat(self):
        # Test that each of several string-accepting methods return pd.NaT
        result1, _ = parsing.parse_datetime_string_with_reso("NaT")
        result2 = to_datetime("NaT")
        result3 = Timestamp("NaT")
        result4 = DatetimeIndex(["NaT"])[0]
        assert result1 is NaT
        assert result2 is NaT
        assert result3 is NaT
        assert result4 is NaT

    @pytest.mark.parametrize(
        "date_str, dayfirst, yearfirst, expected",
        [
            ("10-11-12", False, False, datetime(2012, 10, 11)),
            ("10-11-12", True, False, datetime(2012, 11, 10)),
            ("10-11-12", False, True, datetime(2010, 11, 12)),
            ("10-11-12", True, True, datetime(2010, 12, 11)),
            ("20/12/21", False, False, datetime(2021, 12, 20)),
            ("20/12/21", True, False, datetime(2021, 12, 20)),
            ("20/12/21", False, True, datetime(2020, 12, 21)),
            ("20/12/21", True, True, datetime(2020, 12, 21)),
        ],
    )
    def test_parsers_dayfirst_yearfirst(
        self, cache, date_str, dayfirst, yearfirst, expected
    ):
        # OK
        # 2.5.1 10-11-12   [dayfirst=0, yearfirst=0] -> 2012-10-11 00:00:00
        # 2.5.2 10-11-12   [dayfirst=0, yearfirst=1] -> 2012-10-11 00:00:00
        # 2.5.3 10-11-12   [dayfirst=0, yearfirst=0] -> 2012-10-11 00:00:00

        # OK
        # 2.5.1 10-11-12   [dayfirst=0, yearfirst=1] -> 2010-11-12 00:00:00
        # 2.5.2 10-11-12   [dayfirst=0, yearfirst=1] -> 2010-11-12 00:00:00
        # 2.5.3 10-11-12   [dayfirst=0, yearfirst=1] -> 2010-11-12 00:00:00

        # bug fix in 2.5.2
        # 2.5.1 10-11-12   [dayfirst=1, yearfirst=1] -> 2010-11-12 00:00:00
        # 2.5.2 10-11-12   [dayfirst=1, yearfirst=1] -> 2010-12-11 00:00:00
        # 2.5.3 10-11-12   [dayfirst=1, yearfirst=1] -> 2010-12-11 00:00:00

        # OK
        # 2.5.1 10-11-12   [dayfirst=1, yearfirst=0] -> 2012-11-10 00:00:00
        # 2.5.2 10-11-12   [dayfirst=1, yearfirst=0] -> 2012-11-10 00:00:00
        # 2.5.3 10-11-12   [dayfirst=1, yearfirst=0] -> 2012-11-10 00:00:00

        # OK
        # 2.5.1 20/12/21   [dayfirst=0, yearfirst=0] -> 2021-12-20 00:00:00
        # 2.5.2 20/12/21   [dayfirst=0, yearfirst=0] -> 2021-12-20 00:00:00
        # 2.5.3 20/12/21   [dayfirst=0, yearfirst=0] -> 2021-12-20 00:00:00

        # OK
        # 2.5.1 20/12/21   [dayfirst=0, yearfirst=1] -> 2020-12-21 00:00:00
        # 2.5.2 20/12/21   [dayfirst=0, yearfirst=1] -> 2020-12-21 00:00:00
        # 2.5.3 20/12/21   [dayfirst=0, yearfirst=1] -> 2020-12-21 00:00:00

        # revert of bug in 2.5.2
        # 2.5.1 20/12/21   [dayfirst=1, yearfirst=1] -> 2020-12-21 00:00:00
        # 2.5.2 20/12/21   [dayfirst=1, yearfirst=1] -> month must be in 1..12
        # 2.5.3 20/12/21   [dayfirst=1, yearfirst=1] -> 2020-12-21 00:00:00

        # OK
        # 2.5.1 20/12/21   [dayfirst=1, yearfirst=0] -> 2021-12-20 00:00:00
        # 2.5.2 20/12/21   [dayfirst=1, yearfirst=0] -> 2021-12-20 00:00:00
        # 2.5.3 20/12/21   [dayfirst=1, yearfirst=0] -> 2021-12-20 00:00:00

        # str : dayfirst, yearfirst, expected

        # compare with dateutil result
        dateutil_result = parse(date_str, dayfirst=dayfirst, yearfirst=yearfirst)
        assert dateutil_result == expected

        result1, _ = parsing.parse_datetime_string_with_reso(
            date_str, dayfirst=dayfirst, yearfirst=yearfirst
        )

        # we don't support dayfirst/yearfirst here:
        if not dayfirst and not yearfirst:
            result2 = Timestamp(date_str)
            assert result2 == expected

        result3 = to_datetime(
            date_str, dayfirst=dayfirst, yearfirst=yearfirst, cache=cache
        )

        result4 = DatetimeIndex([date_str], dayfirst=dayfirst, yearfirst=yearfirst)[0]

        assert result1 == expected
        assert result3 == expected
        assert result4 == expected

    @pytest.mark.parametrize(
        "date_str, exp_def",
        [["10:15", datetime(1, 1, 1, 10, 15)], ["9:05", datetime(1, 1, 1, 9, 5)]],
    )
    def test_parsers_timestring(self, date_str, exp_def):
        # must be the same as dateutil result
        exp_now = parse(date_str)

        result1, _ = parsing.parse_datetime_string_with_reso(date_str)
        result2 = to_datetime(date_str)
        result3 = to_datetime([date_str])
        result4 = Timestamp(date_str)
        result5 = DatetimeIndex([date_str])[0]
        # parse time string return time string based on default date
        # others are not, and can't be changed because it is used in
        # time series plot
        assert result1 == exp_def
        assert result2 == exp_now
        assert result3 == exp_now
        assert result4 == exp_now
        assert result5 == exp_now

    @pytest.mark.parametrize(
        "dt_string, tz, dt_string_repr",
        [
            (
                "2013-01-01 05:45+0545",
                timezone(timedelta(minutes=345)),
                "Timestamp('2013-01-01 05:45:00+0545', tz='UTC+05:45')",
            ),
            (
                "2013-01-01 05:30+0530",
                timezone(timedelta(minutes=330)),
                "Timestamp('2013-01-01 05:30:00+0530', tz='UTC+05:30')",
            ),
        ],
    )
    def test_parsers_timezone_minute_offsets_roundtrip(
        self, cache, dt_string, tz, dt_string_repr
    ):
        # GH11708
        base = to_datetime("2013-01-01 00:00:00", cache=cache)
        base = base.tz_localize("UTC").tz_convert(tz)
        dt_time = to_datetime(dt_string, cache=cache)
        assert base == dt_time
        assert dt_string_repr == repr(dt_time)


@pytest.fixture(params=["D", "s", "ms", "us", "ns"])
def units(request):
    """Day and some time units.

    * D
    * s
    * ms
    * us
    * ns
    """
    return request.param


@pytest.fixture
def epoch_1960():
    """Timestamp at 1960-01-01."""
    return Timestamp("1960-01-01")


@pytest.fixture
def units_from_epochs():
    return list(range(5))


@pytest.fixture(params=["timestamp", "pydatetime", "datetime64", "str_1960"])
def epochs(epoch_1960, request):
    """Timestamp at 1960-01-01 in various forms.

    * Timestamp
    * datetime.datetime
    * numpy.datetime64
    * str
    """
    assert request.param in {"timestamp", "pydatetime", "datetime64", "str_1960"}
    if request.param == "timestamp":
        return epoch_1960
    elif request.param == "pydatetime":
        return epoch_1960.to_pydatetime()
    elif request.param == "datetime64":
        return epoch_1960.to_datetime64()
    else:
        return str(epoch_1960)


@pytest.fixture
def julian_dates():
    return date_range("2014-1-1", periods=10).to_julian_date().values


class TestOrigin:
    def test_origin_and_unit(self):
        # GH#42624
        ts = to_datetime(1, unit="s", origin=1)
        expected = Timestamp("1970-01-01 00:00:02")
        assert ts == expected

        ts = to_datetime(1, unit="s", origin=1_000_000_000)
        expected = Timestamp("2001-09-09 01:46:41")
        assert ts == expected

    def test_julian(self, julian_dates):
        # gh-11276, gh-11745
        # for origin as julian

        result = Series(to_datetime(julian_dates, unit="D", origin="julian"))
        expected = Series(
            to_datetime(julian_dates - Timestamp(0).to_julian_date(), unit="D")
        )
        tm.assert_series_equal(result, expected)

    def test_unix(self):
        result = Series(to_datetime([0, 1, 2], unit="D", origin="unix"))
        expected = Series(
            [Timestamp("1970-01-01"), Timestamp("1970-01-02"), Timestamp("1970-01-03")],
            dtype="M8[ns]",
        )
        tm.assert_series_equal(result, expected)

    def test_julian_round_trip(self):
        result = to_datetime(2456658, origin="julian", unit="D")
        assert result.to_julian_date() == 2456658

        # out-of-bounds
        msg = "1 is Out of Bounds for origin='julian'"
        with pytest.raises(ValueError, match=msg):
            to_datetime(1, origin="julian", unit="D")

    def test_invalid_unit(self, units, julian_dates):
        # checking for invalid combination of origin='julian' and unit != D
        if units != "D":
            msg = "unit must be 'D' for origin='julian'"
            with pytest.raises(ValueError, match=msg):
                to_datetime(julian_dates, unit=units, origin="julian")

    @pytest.mark.parametrize("unit", ["ns", "D"])
    def test_invalid_origin(self, unit):
        # need to have a numeric specified
        msg = "it must be numeric with a unit specified"
        with pytest.raises(ValueError, match=msg):
            to_datetime("2005-01-01", origin="1960-01-01", unit=unit)

    @pytest.mark.parametrize(
        "epochs",
        [
            Timestamp(1960, 1, 1),
            datetime(1960, 1, 1),
            "1960-01-01",
            np.datetime64("1960-01-01"),
        ],
    )
    def test_epoch(self, units, epochs):
        epoch_1960 = Timestamp(1960, 1, 1)
        units_from_epochs = np.arange(5, dtype=np.int64)
        expected = Series(
            [pd.Timedelta(x, unit=units) + epoch_1960 for x in units_from_epochs]
        )

        result = Series(to_datetime(units_from_epochs, unit=units, origin=epochs))
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize(
        "origin, exc",
        [
            ("random_string", ValueError),
            ("epoch", ValueError),
            ("13-24-1990", ValueError),
            (datetime(1, 1, 1), OutOfBoundsDatetime),
        ],
    )
    def test_invalid_origins(self, origin, exc, units, units_from_epochs):
        msg = "|".join(
            [
                f"origin {origin} is Out of Bounds",
                f"origin {origin} cannot be converted to a Timestamp",
                "Cannot cast .* to unit='ns' without overflow",
            ]
        )
        with pytest.raises(exc, match=msg):
            to_datetime(units_from_epochs, unit=units, origin=origin)

    def test_invalid_origins_tzinfo(self):
        # GH16842
        with pytest.raises(ValueError, match="must be tz-naive"):
            to_datetime(1, unit="D", origin=datetime(2000, 1, 1, tzinfo=pytz.utc))

    def test_incorrect_value_exception(self):
        # GH47495
        msg = (
            "Unknown datetime string format, unable to parse: yesterday, at position 1"
        )
        with pytest.raises(ValueError, match=msg):
            to_datetime(["today", "yesterday"])

    @pytest.mark.parametrize(
        "format, warning",
        [
            (None, UserWarning),
            ("%Y-%m-%d %H:%M:%S", None),
            ("%Y-%d-%m %H:%M:%S", None),
        ],
    )
    def test_to_datetime_out_of_bounds_with_format_arg(self, format, warning):
        # see gh-23830
        msg = r"^Out of bounds nanosecond timestamp: 2417-10-10 00:00:00, at position 0"
        with pytest.raises(OutOfBoundsDatetime, match=msg):
            to_datetime("2417-10-10 00:00:00", format=format)

    @pytest.mark.parametrize(
        "arg, origin, expected_str",
        [
            [200 * 365, "unix", "2169-11-13 00:00:00"],
            [200 * 365, "1870-01-01", "2069-11-13 00:00:00"],
            [300 * 365, "1870-01-01", "2169-10-20 00:00:00"],
        ],
    )
    def test_processing_order(self, arg, origin, expected_str):
        # make sure we handle out-of-bounds *before*
        # constructing the dates

        result = to_datetime(arg, unit="D", origin=origin)
        expected = Timestamp(expected_str)
        assert result == expected

        result = to_datetime(200 * 365, unit="D", origin="1870-01-01")
        expected = Timestamp("2069-11-13 00:00:00")
        assert result == expected

        result = to_datetime(300 * 365, unit="D", origin="1870-01-01")
        expected = Timestamp("2169-10-20 00:00:00")
        assert result == expected

    @pytest.mark.parametrize(
        "offset,utc,exp",
        [
            ["Z", True, "2019-01-01T00:00:00.000Z"],
            ["Z", None, "2019-01-01T00:00:00.000Z"],
            ["-01:00", True, "2019-01-01T01:00:00.000Z"],
            ["-01:00", None, "2019-01-01T00:00:00.000-01:00"],
        ],
    )
    def test_arg_tz_ns_unit(self, offset, utc, exp):
        # GH 25546
        arg = "2019-01-01T00:00:00.000" + offset
        result = to_datetime([arg], unit="ns", utc=utc)
        expected = to_datetime([exp]).as_unit("ns")
        tm.assert_index_equal(result, expected)


class TestShouldCache:
    @pytest.mark.parametrize(
        "listlike,do_caching",
        [
            ([1, 2, 3, 4, 5, 6, 7, 8, 9, 0], False),
            ([1, 1, 1, 1, 4, 5, 6, 7, 8, 9], True),
        ],
    )
    def test_should_cache(self, listlike, do_caching):
        assert (
            tools.should_cache(listlike, check_count=len(listlike), unique_share=0.7)
            == do_caching
        )

    @pytest.mark.parametrize(
        "unique_share,check_count, err_message",
        [
            (0.5, 11, r"check_count must be in next bounds: \[0; len\(arg\)\]"),
            (10, 2, r"unique_share must be in next bounds: \(0; 1\)"),
        ],
    )
    def test_should_cache_errors(self, unique_share, check_count, err_message):
        arg = [5] * 10

        with pytest.raises(AssertionError, match=err_message):
            tools.should_cache(arg, unique_share, check_count)

    @pytest.mark.parametrize(
        "listlike",
        [
            (deque([Timestamp("2010-06-02 09:30:00")] * 51)),
            ([Timestamp("2010-06-02 09:30:00")] * 51),
            (tuple([Timestamp("2010-06-02 09:30:00")] * 51)),
        ],
    )
    def test_no_slicing_errors_in_should_cache(self, listlike):
        # GH#29403
        assert tools.should_cache(listlike) is True


def test_nullable_integer_to_datetime():
    # Test for #30050
    ser = Series([1, 2, None, 2**61, None])
    ser = ser.astype("Int64")
    ser_copy = ser.copy()

    res = to_datetime(ser, unit="ns")

    expected = Series(
        [
            np.datetime64("1970-01-01 00:00:00.000000001"),
            np.datetime64("1970-01-01 00:00:00.000000002"),
            np.datetime64("NaT"),
            np.datetime64("2043-01-25 23:56:49.213693952"),
            np.datetime64("NaT"),
        ]
    )
    tm.assert_series_equal(res, expected)
    # Check that ser isn't mutated
    tm.assert_series_equal(ser, ser_copy)


@pytest.mark.parametrize("klass", [np.array, list])
def test_na_to_datetime(nulls_fixture, klass):
    if isinstance(nulls_fixture, Decimal):
        with pytest.raises(TypeError, match="not convertible to datetime"):
            to_datetime(klass([nulls_fixture]))

    else:
        result = to_datetime(klass([nulls_fixture]))

        assert result[0] is NaT


@pytest.mark.parametrize("errors", ["raise", "coerce", "ignore"])
@pytest.mark.parametrize(
    "args, format",
    [
        (["03/24/2016", "03/25/2016", ""], "%m/%d/%Y"),
        (["2016-03-24", "2016-03-25", ""], "%Y-%m-%d"),
    ],
    ids=["non-ISO8601", "ISO8601"],
)
def test_empty_string_datetime(errors, args, format):
    # GH13044, GH50251
    td = Series(args)

    # coerce empty string to pd.NaT
    result = to_datetime(td, format=format, errors=errors)
    expected = Series(["2016-03-24", "2016-03-25", NaT], dtype="datetime64[ns]")
    tm.assert_series_equal(expected, result)


def test_empty_string_datetime_coerce__unit():
    # GH13044
    # coerce empty string to pd.NaT
    result = to_datetime([1, ""], unit="s", errors="coerce")
    expected = DatetimeIndex(["1970-01-01 00:00:01", "NaT"], dtype="datetime64[ns]")
    tm.assert_index_equal(expected, result)

    # verify that no exception is raised even when errors='raise' is set
    result = to_datetime([1, ""], unit="s", errors="raise")
    tm.assert_index_equal(expected, result)


@pytest.mark.parametrize("cache", [True, False])
def test_to_datetime_monotonic_increasing_index(cache):
    # GH28238
    cstart = start_caching_at
    times = date_range(Timestamp("1980"), periods=cstart, freq="YS")
    times = times.to_frame(index=False, name="DT").sample(n=cstart, random_state=1)
    times.index = times.index.to_series().astype(float) / 1000
    result = to_datetime(times.iloc[:, 0], cache=cache)
    expected = times.iloc[:, 0]
    tm.assert_series_equal(result, expected)


@pytest.mark.parametrize(
    "series_length",
    [40, start_caching_at, (start_caching_at + 1), (start_caching_at + 5)],
)
def test_to_datetime_cache_coerce_50_lines_outofbounds(series_length):
    # GH#45319
    ser = Series(
        [datetime.fromisoformat("1446-04-12 00:00:00+00:00")]
        + ([datetime.fromisoformat("1991-10-20 00:00:00+00:00")] * series_length),
        dtype=object,
    )
    result1 = to_datetime(ser, errors="coerce", utc=True)

    expected1 = Series(
        [NaT] + ([Timestamp("1991-10-20 00:00:00+00:00")] * series_length)
    )

    tm.assert_series_equal(result1, expected1)

    result2 = to_datetime(ser, errors="ignore", utc=True)

    expected2 = Series(
        [datetime.fromisoformat("1446-04-12 00:00:00+00:00")]
        + ([datetime.fromisoformat("1991-10-20 00:00:00+00:00")] * series_length)
    )

    tm.assert_series_equal(result2, expected2)

    with pytest.raises(OutOfBoundsDatetime, match="Out of bounds nanosecond timestamp"):
        to_datetime(ser, errors="raise", utc=True)


def test_to_datetime_format_f_parse_nanos():
    # GH 48767
    timestamp = "15/02/2020 02:03:04.123456789"
    timestamp_format = "%d/%m/%Y %H:%M:%S.%f"
    result = to_datetime(timestamp, format=timestamp_format)
    expected = Timestamp(
        year=2020,
        month=2,
        day=15,
        hour=2,
        minute=3,
        second=4,
        microsecond=123456,
        nanosecond=789,
    )
    assert result == expected


def test_to_datetime_mixed_iso8601():
    # https://github.com/pandas-dev/pandas/issues/50411
    result = to_datetime(["2020-01-01", "2020-01-01 05:00:00"], format="ISO8601")
    expected = DatetimeIndex(["2020-01-01 00:00:00", "2020-01-01 05:00:00"])
    tm.assert_index_equal(result, expected)


def test_to_datetime_mixed_other():
    # https://github.com/pandas-dev/pandas/issues/50411
    result = to_datetime(["01/11/2000", "12 January 2000"], format="mixed")
    expected = DatetimeIndex(["2000-01-11", "2000-01-12"])
    tm.assert_index_equal(result, expected)


@pytest.mark.parametrize("exact", [True, False])
@pytest.mark.parametrize("format", ["ISO8601", "mixed"])
def test_to_datetime_mixed_or_iso_exact(exact, format):
    msg = "Cannot use 'exact' when 'format' is 'mixed' or 'ISO8601'"
    with pytest.raises(ValueError, match=msg):
        to_datetime(["2020-01-01"], exact=exact, format=format)


def test_to_datetime_mixed_not_necessarily_iso8601_raise():
    # https://github.com/pandas-dev/pandas/issues/50411
    with pytest.raises(
        ValueError, match="Time data 01-01-2000 is not ISO8601 format, at position 1"
    ):
        to_datetime(["2020-01-01", "01-01-2000"], format="ISO8601")


@pytest.mark.parametrize(
    ("errors", "expected"),
    [
        ("coerce", DatetimeIndex(["2020-01-01 00:00:00", NaT])),
        ("ignore", Index(["2020-01-01", "01-01-2000"], dtype="str")),
    ],
)
def test_to_datetime_mixed_not_necessarily_iso8601_coerce(errors, expected):
    # https://github.com/pandas-dev/pandas/issues/50411
    result = to_datetime(["2020-01-01", "01-01-2000"], format="ISO8601", errors=errors)
    tm.assert_index_equal(result, expected)


def test_ignoring_unknown_tz_deprecated():
    # GH#18702, GH#51476
    dtstr = "2014 Jan 9 05:15 FAKE"
    msg = 'un-recognized timezone "FAKE". Dropping unrecognized timezones is deprecated'
    with tm.assert_produces_warning(FutureWarning, match=msg):
        res = Timestamp(dtstr)
    assert res == Timestamp(dtstr[:-5])

    with tm.assert_produces_warning(FutureWarning):
        res = to_datetime(dtstr)
    assert res == to_datetime(dtstr[:-5])
    with tm.assert_produces_warning(FutureWarning):
        res = to_datetime([dtstr])
    tm.assert_index_equal(res, to_datetime([dtstr[:-5]]))


def test_from_numeric_arrow_dtype(any_numeric_ea_dtype):
    # GH 52425
    pytest.importorskip("pyarrow")
    ser = Series([1, 2], dtype=f"{any_numeric_ea_dtype.lower()}[pyarrow]")
    result = to_datetime(ser)
    expected = Series([1, 2], dtype="datetime64[ns]")
    tm.assert_series_equal(result, expected)


def test_to_datetime_with_empty_str_utc_false_format_mixed():
    # GH 50887
    vals = ["2020-01-01 00:00+00:00", ""]
    result = to_datetime(vals, format="mixed")
    expected = Index([Timestamp("2020-01-01 00:00+00:00"), "NaT"], dtype="M8[ns, UTC]")
    tm.assert_index_equal(result, expected)

    # Check that a couple of other similar paths work the same way
    alt = to_datetime(vals)
    tm.assert_index_equal(alt, expected)
    alt2 = DatetimeIndex(vals)
    tm.assert_index_equal(alt2, expected)


def test_to_datetime_with_empty_str_utc_false_offsets_and_format_mixed():
    # GH 50887
    msg = "parsing datetimes with mixed time zones will raise an error"

    with tm.assert_produces_warning(FutureWarning, match=msg):
        to_datetime(
            ["2020-01-01 00:00+00:00", "2020-01-01 00:00+02:00", ""], format="mixed"
        )


def test_to_datetime_mixed_tzs_mixed_types():
    # GH#55793, GH#55693 mismatched tzs but one is str and other is
    #  datetime object
    ts = Timestamp("2016-01-02 03:04:05", tz="US/Pacific")
    dtstr = "2023-10-30 15:06+01"
    arr = [ts, dtstr]

    msg = (
        "Mixed timezones detected. pass utc=True in to_datetime or tz='UTC' "
        "in DatetimeIndex to convert to a common timezone"
    )
    with pytest.raises(ValueError, match=msg):
        to_datetime(arr)
    with pytest.raises(ValueError, match=msg):
        to_datetime(arr, format="mixed")
    with pytest.raises(ValueError, match=msg):
        DatetimeIndex(arr)


def test_to_datetime_mixed_types_matching_tzs():
    # GH#55793
    dtstr = "2023-11-01 09:22:03-07:00"
    ts = Timestamp(dtstr)
    arr = [ts, dtstr]
    res1 = to_datetime(arr)
    res2 = to_datetime(arr[::-1])[::-1]
    res3 = to_datetime(arr, format="mixed")
    res4 = DatetimeIndex(arr)

    expected = DatetimeIndex([ts, ts])
    tm.assert_index_equal(res1, expected)
    tm.assert_index_equal(res2, expected)
    tm.assert_index_equal(res3, expected)
    tm.assert_index_equal(res4, expected)


dtstr = "2020-01-01 00:00+00:00"
ts = Timestamp(dtstr)


@pytest.mark.filterwarnings("ignore:Could not infer format:UserWarning")
@pytest.mark.parametrize(
    "aware_val",
    [dtstr, Timestamp(dtstr)],
    ids=lambda x: type(x).__name__,
)
@pytest.mark.parametrize(
    "naive_val",
    [dtstr[:-6], ts.tz_localize(None), ts.date(), ts.asm8, ts.value, float(ts.value)],
    ids=lambda x: type(x).__name__,
)
@pytest.mark.parametrize("naive_first", [True, False])
def test_to_datetime_mixed_awareness_mixed_types(aware_val, naive_val, naive_first):
    # GH#55793, GH#55693
    # Empty string parses to NaT
    vals = [aware_val, naive_val, ""]

    vec = vals
    if naive_first:
        # alas, the behavior is order-dependent, so we test both ways
        vec = [naive_val, aware_val, ""]

    # both_strs-> paths that were previously already deprecated with warning
    #  issued in _array_to_datetime_object
    both_strs = isinstance(aware_val, str) and isinstance(naive_val, str)
    has_numeric = isinstance(naive_val, (int, float))

    depr_msg = "In a future version of pandas, parsing datetimes with mixed time zones"

    first_non_null = next(x for x in vec if x != "")
    # if first_non_null is a not a string, _guess_datetime_format_for_array
    #  doesn't guess a format so we don't go through array_strptime
    if not isinstance(first_non_null, str):
        # that case goes through array_strptime which has different behavior
        msg = "Cannot mix tz-aware with tz-naive values"
        if naive_first and isinstance(aware_val, Timestamp):
            if isinstance(naive_val, Timestamp):
                msg = "Tz-aware datetime.datetime cannot be converted to datetime64"
            with pytest.raises(ValueError, match=msg):
                to_datetime(vec)
        else:
            with pytest.raises(ValueError, match=msg):
                to_datetime(vec)

        # No warning/error with utc=True
        to_datetime(vec, utc=True)

    elif has_numeric and vec.index(aware_val) < vec.index(naive_val):
        msg = "time data .* doesn't match format"
        with pytest.raises(ValueError, match=msg):
            to_datetime(vec)
        with pytest.raises(ValueError, match=msg):
            to_datetime(vec, utc=True)

    elif both_strs and vec.index(aware_val) < vec.index(naive_val):
        msg = r"time data \"2020-01-01 00:00\" doesn't match format"
        with pytest.raises(ValueError, match=msg):
            to_datetime(vec)
        with pytest.raises(ValueError, match=msg):
            to_datetime(vec, utc=True)

    elif both_strs and vec.index(naive_val) < vec.index(aware_val):
        msg = "unconverted data remains when parsing with format"
        with pytest.raises(ValueError, match=msg):
            to_datetime(vec)
        with pytest.raises(ValueError, match=msg):
            to_datetime(vec, utc=True)

    else:
        with tm.assert_produces_warning(FutureWarning, match=depr_msg):
            to_datetime(vec)

        # No warning/error with utc=True
        to_datetime(vec, utc=True)

    if both_strs:
        with tm.assert_produces_warning(FutureWarning, match=depr_msg):
            to_datetime(vec, format="mixed")
        with tm.assert_produces_warning(FutureWarning, match=depr_msg):
            msg = "DatetimeIndex has mixed timezones"
            with pytest.raises(TypeError, match=msg):
                DatetimeIndex(vec)
    else:
        msg = "Cannot mix tz-aware with tz-naive values"
        if naive_first and isinstance(aware_val, Timestamp):
            if isinstance(naive_val, Timestamp):
                msg = "Tz-aware datetime.datetime cannot be converted to datetime64"
            with pytest.raises(ValueError, match=msg):
                to_datetime(vec, format="mixed")
            with pytest.raises(ValueError, match=msg):
                DatetimeIndex(vec)
        else:
            with pytest.raises(ValueError, match=msg):
                to_datetime(vec, format="mixed")
            with pytest.raises(ValueError, match=msg):
                DatetimeIndex(vec)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\cffi\pkgconfig.py ===
# pkg-config, https://www.freedesktop.org/wiki/Software/pkg-config/ integration for cffi
import sys, os, subprocess

from .error import PkgConfigError


def merge_flags(cfg1, cfg2):
    """Merge values from cffi config flags cfg2 to cf1

    Example:
        merge_flags({"libraries": ["one"]}, {"libraries": ["two"]})
        {"libraries": ["one", "two"]}
    """
    for key, value in cfg2.items():
        if key not in cfg1:
            cfg1[key] = value
        else:
            if not isinstance(cfg1[key], list):
                raise TypeError("cfg1[%r] should be a list of strings" % (key,))
            if not isinstance(value, list):
                raise TypeError("cfg2[%r] should be a list of strings" % (key,))
            cfg1[key].extend(value)
    return cfg1


def call(libname, flag, encoding=sys.getfilesystemencoding()):
    """Calls pkg-config and returns the output if found
    """
    a = ["pkg-config", "--print-errors"]
    a.append(flag)
    a.append(libname)
    try:
        pc = subprocess.Popen(a, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    except EnvironmentError as e:
        raise PkgConfigError("cannot run pkg-config: %s" % (str(e).strip(),))

    bout, berr = pc.communicate()
    if pc.returncode != 0:
        try:
            berr = berr.decode(encoding)
        except Exception:
            pass
        raise PkgConfigError(berr.strip())

    if sys.version_info >= (3,) and not isinstance(bout, str):   # Python 3.x
        try:
            bout = bout.decode(encoding)
        except UnicodeDecodeError:
            raise PkgConfigError("pkg-config %s %s returned bytes that cannot "
                                 "be decoded with encoding %r:\n%r" %
                                 (flag, libname, encoding, bout))

    if os.altsep != '\\' and '\\' in bout:
        raise PkgConfigError("pkg-config %s %s returned an unsupported "
                             "backslash-escaped output:\n%r" %
                             (flag, libname, bout))
    return bout


def flags_from_pkgconfig(libs):
    r"""Return compiler line flags for FFI.set_source based on pkg-config output

    Usage
        ...
        ffibuilder.set_source("_foo", pkgconfig = ["libfoo", "libbar >= 1.8.3"])

    If pkg-config is installed on build machine, then arguments include_dirs,
    library_dirs, libraries, define_macros, extra_compile_args and
    extra_link_args are extended with an output of pkg-config for libfoo and
    libbar.

    Raises PkgConfigError in case the pkg-config call fails.
    """

    def get_include_dirs(string):
        return [x[2:] for x in string.split() if x.startswith("-I")]

    def get_library_dirs(string):
        return [x[2:] for x in string.split() if x.startswith("-L")]

    def get_libraries(string):
        return [x[2:] for x in string.split() if x.startswith("-l")]

    # convert -Dfoo=bar to list of tuples [("foo", "bar")] expected by distutils
    def get_macros(string):
        def _macro(x):
            x = x[2:]    # drop "-D"
            if '=' in x:
                return tuple(x.split("=", 1))  # "-Dfoo=bar" => ("foo", "bar")
            else:
                return (x, None)               # "-Dfoo" => ("foo", None)
        return [_macro(x) for x in string.split() if x.startswith("-D")]

    def get_other_cflags(string):
        return [x for x in string.split() if not x.startswith("-I") and
                                             not x.startswith("-D")]

    def get_other_libs(string):
        return [x for x in string.split() if not x.startswith("-L") and
                                             not x.startswith("-l")]

    # return kwargs for given libname
    def kwargs(libname):
        fse = sys.getfilesystemencoding()
        all_cflags = call(libname, "--cflags")
        all_libs = call(libname, "--libs")
        return {
            "include_dirs": get_include_dirs(all_cflags),
            "library_dirs": get_library_dirs(all_libs),
            "libraries": get_libraries(all_libs),
            "define_macros": get_macros(all_cflags),
            "extra_compile_args": get_other_cflags(all_cflags),
            "extra_link_args": get_other_libs(all_libs),
            }

    # merge all arguments together
    ret = {}
    for libname in libs:
        lib_flags = kwargs(libname)
        merge_flags(ret, lib_flags)
    return ret

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\colorama\initialise.py ===
# Copyright Jonathan Hartley 2013. BSD 3-Clause license, see LICENSE file.
import atexit
import contextlib
import sys

from .ansitowin32 import AnsiToWin32


def _wipe_internal_state_for_tests():
    global orig_stdout, orig_stderr
    orig_stdout = None
    orig_stderr = None

    global wrapped_stdout, wrapped_stderr
    wrapped_stdout = None
    wrapped_stderr = None

    global atexit_done
    atexit_done = False

    global fixed_windows_console
    fixed_windows_console = False

    try:
        # no-op if it wasn't registered
        atexit.unregister(reset_all)
    except AttributeError:
        # python 2: no atexit.unregister. Oh well, we did our best.
        pass


def reset_all():
    if AnsiToWin32 is not None:    # Issue #74: objects might become None at exit
        AnsiToWin32(orig_stdout).reset_all()


def init(autoreset=False, convert=None, strip=None, wrap=True):

    if not wrap and any([autoreset, convert, strip]):
        raise ValueError('wrap=False conflicts with any other arg=True')

    global wrapped_stdout, wrapped_stderr
    global orig_stdout, orig_stderr

    orig_stdout = sys.stdout
    orig_stderr = sys.stderr

    if sys.stdout is None:
        wrapped_stdout = None
    else:
        sys.stdout = wrapped_stdout = \
            wrap_stream(orig_stdout, convert, strip, autoreset, wrap)
    if sys.stderr is None:
        wrapped_stderr = None
    else:
        sys.stderr = wrapped_stderr = \
            wrap_stream(orig_stderr, convert, strip, autoreset, wrap)

    global atexit_done
    if not atexit_done:
        atexit.register(reset_all)
        atexit_done = True


def deinit():
    if orig_stdout is not None:
        sys.stdout = orig_stdout
    if orig_stderr is not None:
        sys.stderr = orig_stderr


def just_fix_windows_console():
    global fixed_windows_console

    if sys.platform != "win32":
        return
    if fixed_windows_console:
        return
    if wrapped_stdout is not None or wrapped_stderr is not None:
        # Someone already ran init() and it did stuff, so we won't second-guess them
        return

    # On newer versions of Windows, AnsiToWin32.__init__ will implicitly enable the
    # native ANSI support in the console as a side-effect. We only need to actually
    # replace sys.stdout/stderr if we're in the old-style conversion mode.
    new_stdout = AnsiToWin32(sys.stdout, convert=None, strip=None, autoreset=False)
    if new_stdout.convert:
        sys.stdout = new_stdout
    new_stderr = AnsiToWin32(sys.stderr, convert=None, strip=None, autoreset=False)
    if new_stderr.convert:
        sys.stderr = new_stderr

    fixed_windows_console = True

@contextlib.contextmanager
def colorama_text(*args, **kwargs):
    init(*args, **kwargs)
    try:
        yield
    finally:
        deinit()


def reinit():
    if wrapped_stdout is not None:
        sys.stdout = wrapped_stdout
    if wrapped_stderr is not None:
        sys.stderr = wrapped_stderr


def wrap_stream(stream, convert, strip, autoreset, wrap):
    if wrap:
        wrapper = AnsiToWin32(stream,
            convert=convert, strip=strip, autoreset=autoreset)
        if wrapper.should_wrap():
            stream = wrapper.stream
    return stream


# Use this for initial setup as well, to reduce code duplication
_wipe_internal_state_for_tests()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\quantization\operators\embed_layernorm.py ===
import logging

import onnx
from onnx import onnx_pb as onnx_proto  # noqa: F401

from ..quant_utils import attribute_to_kwarg, ms_domain
from .base_operator import QuantOperatorBase

"""
Quantizes the EmbedLayerNorm fused ONNXRuntime Op.

This Quant operator keeps the input and segment IDs at int32 but will quantize all initializer and
weight inputs associated with the node to uint8.
"""


class EmbedLayerNormalizationQuant(QuantOperatorBase):
    def __init__(self, onnx_quantizer, onnx_node):
        super().__init__(onnx_quantizer, onnx_node)

    def should_quantize(self):
        return self.quantizer.should_quantize_node(self.node)

    def quantize(self):
        node = self.node
        assert node.op_type == "EmbedLayerNormalization"

        if len(node.output) > 2:
            logging.info(f"Quantization is not applied to {node.name} since it has 3 outputs")
            return super().quantize()

        """
        Pre-quantization EmbedLayerNorm inputs:
        [0] input_ids (int32)
        [1] segment_ids (int32)
        [2] word_embedding (float32)
        [3] position_embedding (float32)
        [4] segment_embedding (float32)
        [5] gamma (float32)
        [6] beta (float32)
        [7] mask (int32) (optional)
        """
        (
            quantized_input_names,
            zero_point_names,
            scale_names,
            nodes,
        ) = self.quantizer.quantize_activation(node, [2, 3, 4, 5, 6])
        if quantized_input_names is None:
            return super().quantize()

        qembed_layer_norm_name = "" if not node.name else node.name + "_quant"

        """
        Quantized Input Tensor List
        [0] input_ids (int32)
        [1] segment_ids (int32)
        [2] word_embedding (uint8)
        [3] position_embedding (uint8)
        [4] segment_embedding (uint8)
        [5] gamma (uint8)
        [6] beta (uint8)
        [7] mask (int32) (optional)
        [8] word_embedding_scale (float)
        [9] position_embedding_scale (float)
        [10] segment_embedding_scale (float)
        [11] gamma_scale (float)
        [12] beta_scale (float)
        [13] word_embedding_zero_point (uint8)
        [14] position_embedding_zero_point (uint8)
        [15] segment_embedding_zero_point (uint8)
        [16] gamma_zero_point (uint8)
        [17] beta_zero_point (uint8)
        """
        inputs = []
        # 'input_ids'
        inputs.extend([node.input[0]])
        # 'segment_ids'
        inputs.extend([node.input[1]])
        # 'word_embedding_quant'
        inputs.extend([quantized_input_names[0]])
        # 'position_embedding_quant'
        inputs.extend([quantized_input_names[1]])
        # 'segment_embedding_quant'
        inputs.extend([quantized_input_names[2]])
        # 'gamma_quant'
        inputs.extend([quantized_input_names[3]])
        # 'beta_quant'
        inputs.extend([quantized_input_names[4]])
        # 'mask' (optional)
        inputs.extend([node.input[7] if len(node.input) > 7 else ""])

        # Add all scales:
        inputs.extend([scale_names[0]])
        inputs.extend([scale_names[1]])
        inputs.extend([scale_names[2]])
        inputs.extend([scale_names[3]])
        inputs.extend([scale_names[4]])

        # Add all zero points:
        inputs.extend([zero_point_names[0]])
        inputs.extend([zero_point_names[1]])
        inputs.extend([zero_point_names[2]])
        inputs.extend([zero_point_names[3]])
        inputs.extend([zero_point_names[4]])

        kwargs = {}
        for attribute in node.attribute:
            kwargs.update(attribute_to_kwarg(attribute))
        kwargs["domain"] = ms_domain

        qembed_layer_norm_node = onnx.helper.make_node(
            "QEmbedLayerNormalization",
            inputs,
            node.output,
            qembed_layer_norm_name,
            **kwargs,
        )
        nodes.append(qembed_layer_norm_node)

        self.quantizer.new_nodes += nodes

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\quantization\operators\lstm.py ===
import numpy
import onnx
from onnx import onnx_pb as onnx_proto

from ..quant_utils import QuantType, attribute_to_kwarg, ms_domain  # noqa: F401
from .base_operator import QuantOperatorBase

"""
    Quantize LSTM
"""


class LSTMQuant(QuantOperatorBase):
    def __init__(self, onnx_quantizer, onnx_node):
        super().__init__(onnx_quantizer, onnx_node)

    def quantize(self):
        """
        parameter node: LSTM node.
        parameter new_nodes_list: List of new nodes created before processing this node.
        return: a list of nodes in topological order that represents quantized Attention node.
        """
        node = self.node
        assert node.op_type == "LSTM"

        if not self.quantizer.is_valid_quantize_weight(node.input[1]) or not self.quantizer.is_valid_quantize_weight(
            node.input[2]
        ):
            super().quantize()
            return

        model = self.quantizer.model
        W = model.get_initializer(node.input[1])  # noqa: N806
        R = model.get_initializer(node.input[2])  # noqa: N806

        if len(W.dims) != 3 or len(R.dims) != 3:
            super().quantize()
            return

        [W_num_dir, W_4_hidden_size, W_input_size] = W.dims  # noqa: N806
        [R_num_dir, R_4_hidden_size, R_hidden_size] = R.dims  # noqa: N806

        if self.quantizer.is_per_channel():
            del W.dims[0]
            del R.dims[0]
            W.dims[0] = W_num_dir * W_4_hidden_size
            R.dims[0] = R_num_dir * R_4_hidden_size

        quant_input_weight_tuple = self.quantizer.quantize_weight_per_channel(
            node.input[1],
            onnx_proto.TensorProto.INT8,
            0,  # self.quantizer.weight_qType?
        )
        quant_recurrent_weight_tuple = self.quantizer.quantize_weight_per_channel(
            node.input[2],
            onnx_proto.TensorProto.INT8,
            0,  # self.quantizer.weight_qType?
        )

        W_quant_weight = model.get_initializer(quant_input_weight_tuple[0])  # noqa: N806
        R_quant_weight = model.get_initializer(quant_recurrent_weight_tuple[0])  # noqa: N806

        W_quant_array = onnx.numpy_helper.to_array(W_quant_weight)  # noqa: N806
        R_quant_array = onnx.numpy_helper.to_array(R_quant_weight)  # noqa: N806

        W_quant_array = numpy.reshape(W_quant_array, (W_num_dir, W_4_hidden_size, W_input_size))  # noqa: N806
        R_quant_array = numpy.reshape(R_quant_array, (R_num_dir, R_4_hidden_size, R_hidden_size))  # noqa: N806

        W_quant_array = numpy.transpose(W_quant_array, (0, 2, 1))  # noqa: N806
        R_quant_array = numpy.transpose(R_quant_array, (0, 2, 1))  # noqa: N806

        W_quant_tranposed = onnx.numpy_helper.from_array(W_quant_array, quant_input_weight_tuple[0])  # noqa: N806
        R_quant_tranposed = onnx.numpy_helper.from_array(R_quant_array, quant_recurrent_weight_tuple[0])  # noqa: N806

        model.remove_initializers([W_quant_weight, R_quant_weight])
        model.add_initializer(W_quant_tranposed)
        model.add_initializer(R_quant_tranposed)

        W_quant_zp = model.get_initializer(quant_input_weight_tuple[1])  # noqa: N806
        R_quant_zp = model.get_initializer(quant_recurrent_weight_tuple[1])  # noqa: N806
        W_quant_scale = model.get_initializer(quant_input_weight_tuple[2])  # noqa: N806
        R_quant_scale = model.get_initializer(quant_recurrent_weight_tuple[2])  # noqa: N806

        if self.quantizer.is_per_channel():
            W_quant_zp.dims[:] = [W_num_dir, W_4_hidden_size]
            R_quant_zp.dims[:] = [R_num_dir, R_4_hidden_size]
            W_quant_scale.dims[:] = [W_num_dir, W_4_hidden_size]
            R_quant_scale.dims[:] = [R_num_dir, R_4_hidden_size]

        inputs = []
        input_len = len(node.input)
        inputs.extend([node.input[0]])
        inputs.extend([quant_input_weight_tuple[0], quant_recurrent_weight_tuple[0]])
        inputs.extend([node.input[3] if input_len > 3 else ""])
        inputs.extend([node.input[4] if input_len > 4 else ""])
        inputs.extend([node.input[5] if input_len > 5 else ""])
        inputs.extend([node.input[6] if input_len > 6 else ""])
        inputs.extend([node.input[7] if input_len > 7 else ""])
        inputs.extend(
            [
                quant_input_weight_tuple[2],
                quant_input_weight_tuple[1],
                quant_recurrent_weight_tuple[2],
                quant_recurrent_weight_tuple[1],
            ]
        )

        kwargs = {}
        for attribute in node.attribute:
            if attribute.name == "layout":
                continue
            kwargs.update(attribute_to_kwarg(attribute))
        kwargs["domain"] = ms_domain

        quant_lstm_name = "" if not node.name else node.name + "_quant"
        quant_lstm_node = onnx.helper.make_node("DynamicQuantizeLSTM", inputs, node.output, quant_lstm_name, **kwargs)
        self.quantizer.new_nodes.append(quant_lstm_node)

        dequantize_node = self.quantizer._dequantize_value(node.input[0])
        if dequantize_node is not None:
            self.quantizer.new_nodes.append(dequantize_node)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_gemmfastgelu.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

from logging import getLogger

from fusion_base import Fusion
from fusion_utils import NumpyHelper
from onnx import NodeProto, TensorProto, helper
from onnx_model import OnnxModel

logger = getLogger(__name__)


class FusionGemmFastGelu(Fusion):
    def __init__(self, model: OnnxModel):
        super().__init__(model, "GemmFastGelu", "FastGelu", "GemmFastGelu")
        self.shape_infer = None
        self.shape_infer_done = False

    def get_dimensions_from_tensor_proto(self, tensor_proto: TensorProto) -> int | None:
        if tensor_proto.type.tensor_type.HasField("shape"):
            return len(tensor_proto.type.tensor_type.shape.dim)
        else:
            return None

    def get_dimensions(self, input_name: str) -> int | None:
        graph_input = self.model.find_graph_input(input_name)
        if graph_input:
            return self.get_dimensions_from_tensor_proto(graph_input)

        if not self.shape_infer_done:
            self.shape_infer = self.model.infer_runtime_shape(update=True)
            self.shape_infer_done = True

        if self.shape_infer is not None:
            return self.get_dimensions_from_tensor_proto(self.shape_infer.known_vi_[input_name])

        return None

    def fuse(
        self,
        node: NodeProto,
        input_name_to_nodes: dict[str, list[NodeProto]],
        output_name_to_node: dict[str, NodeProto],
    ):
        """
        This pattern is from PyTorch bert model
        Fuse MatMul with FastGelu into one node:

            [root] --> MatMul --> FastGelu -->

        """
        has_bias = False
        if len(node.input) == 2:
            has_bias = True

        match_nodes = self.model.match_parent_path(node, ["MatMul"], [0])
        if match_nodes is None:
            return
        matmul = match_nodes[0]

        # matmul input X should >= two dimension, input weight should be two dimension
        weight_index = -1
        x_dims = 0
        weight = None

        for i, input in enumerate(matmul.input):
            initializer = self.model.get_initializer(input)
            if initializer is None:
                x_dims = self.get_dimensions(matmul.input[i])
            else:
                weight_index = i
                weight = NumpyHelper.to_array(initializer)
        if weight is None:
            return
        if len(weight.shape) != 2:
            return
        if x_dims < len(weight.shape):
            return

        # bias weight should be one dimension
        bias_index = -1
        if has_bias:
            bias_weight = None
            for i, input in enumerate(node.input):
                initializer = self.model.get_initializer(input)
                if initializer is None:
                    continue
                bias_index = i
                bias_weight = NumpyHelper.to_array(initializer)
                break
            if bias_weight is None:
                return
            if len(bias_weight.shape) != 1:
                return

        subgraph_nodes = [node, matmul]
        if not self.model.is_safe_to_fuse_nodes(
            subgraph_nodes, [node.output[0]], input_name_to_nodes, output_name_to_node
        ):
            return

        self.nodes_to_remove.extend(subgraph_nodes)

        inputs = (
            [matmul.input[1 - weight_index], matmul.input[weight_index], node.input[bias_index]]
            if has_bias
            else [matmul.input[1 - weight_index], matmul.input[weight_index]]
        )

        fused_node = helper.make_node(
            "GemmFastGelu",
            inputs=inputs,
            outputs=node.output,
            name=self.model.create_node_name("GemmFastGelu"),
        )
        fused_node.domain = "com.microsoft"
        self.nodes_to_add.append(fused_node)
        self.node_name_to_graph_name[fused_node.name] = self.this_graph_name

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\shape_infer_helper.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

import logging
import os
import sys

# In ORT Package the symbolic_shape_infer.py is in ../tools
file_path = os.path.dirname(__file__)
if os.path.exists(os.path.join(file_path, "../tools/symbolic_shape_infer.py")):
    sys.path.append(os.path.join(file_path, "../tools"))
else:
    sys.path.append(os.path.join(file_path, ".."))

from symbolic_shape_infer import SymbolicShapeInference, get_shape_from_type_proto, sympy  # noqa: E402

logger = logging.getLogger(__name__)


class SymbolicShapeInferenceHelper(SymbolicShapeInference):
    def __init__(self, model, verbose=0, int_max=2**31 - 1, auto_merge=True, guess_output_rank=False):
        super().__init__(int_max, auto_merge, guess_output_rank, verbose)
        self.model_ = model
        self.all_shapes_inferred_: bool = False
        self.is_inferred_: bool = False
        self.dynamic_axis_mapping_: dict[str, int] = {}

    def infer(self, dynamic_axis_mapping: dict[str, int], max_runs: int = 200):
        """Run shape inference, and try replace dynamic axis from string to integer when mapping is provided.

        Args:
            dynamic_axis_mapping (_type_): a dictionary with name of dynamic axis as key, like {"batch_size" : 4}
            max_runs (int, optional): limit maximum number of runs to avoid infinite loop. Defaults to 200.

        Returns:
            bool: whether all shapes has been inferred or not.
        """
        assert dynamic_axis_mapping is not None

        if self.is_inferred_ and self.dynamic_axis_mapping_ == dynamic_axis_mapping:
            return self.all_shapes_inferred_

        self.dynamic_axis_mapping_ = dynamic_axis_mapping

        self._preprocess(self.model_)

        count = 0
        while self.run_:
            logger.debug(f"shape infer run {count}")
            self.all_shapes_inferred_ = self._infer_impl()
            count += 1
            if max_runs > 0 and count >= max_runs:
                break

        self.is_inferred_ = True
        return self.all_shapes_inferred_

    def _get_sympy_shape(self, node, idx):
        """Override it to ensure shape inference by giving the actual value of dynamic axis."""
        sympy_shape = []

        shape = self._get_shape(node, idx)
        if shape:
            for dim in shape:
                if isinstance(dim, str):
                    if dim in self.dynamic_axis_mapping_:
                        sympy_shape.append(self.dynamic_axis_mapping_[dim])
                    elif dim in self.symbolic_dims_:
                        sympy_shape.append(self.symbolic_dims_[dim])
                    else:
                        sympy_shape.append(sympy.Symbol(dim, integer=True))
                else:
                    assert dim is not None
                    sympy_shape.append(dim)
        return sympy_shape

    def get_edge_shape(self, edge):
        """Get shape of an edge.

        Args:
            edge (str): name of edge

        Returns:
            Optional[List[int]]: the shape, or None if shape is unknown
        """
        assert self.all_shapes_inferred_
        if edge not in self.known_vi_:
            print("Cannot retrieve the shape of " + str(edge))
            return None

        type_proto = self.known_vi_[edge].type
        shape = get_shape_from_type_proto(type_proto)

        if shape is not None:
            for i, dim in enumerate(shape):
                if isinstance(dim, str) and dim in self.dynamic_axis_mapping_:
                    shape[i] = self.dynamic_axis_mapping_[dim]

        return shape

    def compare_shape(self, edge, edge_other):
        """Compare shape of two edges.

        Args:
            edge (str): name of edge
            edge_other (str): name of another edge

        Raises:
            Exception: At least one shape is missed for edges to compare

        Returns:
            bool: whether the shape is same or not
        """
        assert self.all_shapes_inferred_
        shape = self.get_edge_shape(edge)
        shape_other = self.get_edge_shape(edge_other)
        if shape is None or shape_other is None:
            raise Exception("At least one shape is missed for edges to compare")
        return shape == shape_other

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\groupby\base.py ===
"""
Provide basic components for groupby.
"""
from __future__ import annotations

import dataclasses
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from collections.abc import Hashable


@dataclasses.dataclass(order=True, frozen=True)
class OutputKey:
    label: Hashable
    position: int


# special case to prevent duplicate plots when catching exceptions when
# forwarding methods from NDFrames
plotting_methods = frozenset(["plot", "hist"])

# cythonized transformations or canned "agg+broadcast", which do not
# require postprocessing of the result by transform.
cythonized_kernels = frozenset(["cumprod", "cumsum", "shift", "cummin", "cummax"])

# List of aggregation/reduction functions.
# These map each group to a single numeric value
reduction_kernels = frozenset(
    [
        "all",
        "any",
        "corrwith",
        "count",
        "first",
        "idxmax",
        "idxmin",
        "last",
        "max",
        "mean",
        "median",
        "min",
        "nunique",
        "prod",
        # as long as `quantile`'s signature accepts only
        # a single quantile value, it's a reduction.
        # GH#27526 might change that.
        "quantile",
        "sem",
        "size",
        "skew",
        "std",
        "sum",
        "var",
    ]
)

# List of transformation functions.
# a transformation is a function that, for each group,
# produces a result that has the same shape as the group.


transformation_kernels = frozenset(
    [
        "bfill",
        "cumcount",
        "cummax",
        "cummin",
        "cumprod",
        "cumsum",
        "diff",
        "ffill",
        "fillna",
        "ngroup",
        "pct_change",
        "rank",
        "shift",
    ]
)

# these are all the public methods on Grouper which don't belong
# in either of the above lists
groupby_other_methods = frozenset(
    [
        "agg",
        "aggregate",
        "apply",
        "boxplot",
        # corr and cov return ngroups*ncolumns rows, so they
        # are neither a transformation nor a reduction
        "corr",
        "cov",
        "describe",
        "dtypes",
        "expanding",
        "ewm",
        "filter",
        "get_group",
        "groups",
        "head",
        "hist",
        "indices",
        "ndim",
        "ngroups",
        "nth",
        "ohlc",
        "pipe",
        "plot",
        "resample",
        "rolling",
        "tail",
        "take",
        "transform",
        "sample",
        "value_counts",
    ]
)
# Valid values  of `name` for `groupby.transform(name)`
# NOTE: do NOT edit this directly. New additions should be inserted
# into the appropriate list above.
transform_kernel_allowlist = reduction_kernels | transformation_kernels

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\io\excel\_calamine.py ===
from __future__ import annotations

from datetime import (
    date,
    datetime,
    time,
    timedelta,
)
from typing import (
    TYPE_CHECKING,
    Any,
    Union,
)

from pandas.compat._optional import import_optional_dependency
from pandas.util._decorators import doc

import pandas as pd
from pandas.core.shared_docs import _shared_docs

from pandas.io.excel._base import BaseExcelReader

if TYPE_CHECKING:
    from python_calamine import (
        CalamineSheet,
        CalamineWorkbook,
    )

    from pandas._typing import (
        FilePath,
        NaTType,
        ReadBuffer,
        Scalar,
        StorageOptions,
    )

_CellValue = Union[int, float, str, bool, time, date, datetime, timedelta]


class CalamineReader(BaseExcelReader["CalamineWorkbook"]):
    @doc(storage_options=_shared_docs["storage_options"])
    def __init__(
        self,
        filepath_or_buffer: FilePath | ReadBuffer[bytes],
        storage_options: StorageOptions | None = None,
        engine_kwargs: dict | None = None,
    ) -> None:
        """
        Reader using calamine engine (xlsx/xls/xlsb/ods).

        Parameters
        ----------
        filepath_or_buffer : str, path to be parsed or
            an open readable stream.
        {storage_options}
        engine_kwargs : dict, optional
            Arbitrary keyword arguments passed to excel engine.
        """
        import_optional_dependency("python_calamine")
        super().__init__(
            filepath_or_buffer,
            storage_options=storage_options,
            engine_kwargs=engine_kwargs,
        )

    @property
    def _workbook_class(self) -> type[CalamineWorkbook]:
        from python_calamine import CalamineWorkbook

        return CalamineWorkbook

    def load_workbook(
        self, filepath_or_buffer: FilePath | ReadBuffer[bytes], engine_kwargs: Any
    ) -> CalamineWorkbook:
        from python_calamine import load_workbook

        return load_workbook(filepath_or_buffer, **engine_kwargs)

    @property
    def sheet_names(self) -> list[str]:
        from python_calamine import SheetTypeEnum

        return [
            sheet.name
            for sheet in self.book.sheets_metadata
            if sheet.typ == SheetTypeEnum.WorkSheet
        ]

    def get_sheet_by_name(self, name: str) -> CalamineSheet:
        self.raise_if_bad_sheet_by_name(name)
        return self.book.get_sheet_by_name(name)

    def get_sheet_by_index(self, index: int) -> CalamineSheet:
        self.raise_if_bad_sheet_by_index(index)
        return self.book.get_sheet_by_index(index)

    def get_sheet_data(
        self, sheet: CalamineSheet, file_rows_needed: int | None = None
    ) -> list[list[Scalar | NaTType | time]]:
        def _convert_cell(value: _CellValue) -> Scalar | NaTType | time:
            if isinstance(value, float):
                val = int(value)
                if val == value:
                    return val
                else:
                    return value
            elif isinstance(value, date):
                return pd.Timestamp(value)
            elif isinstance(value, timedelta):
                return pd.Timedelta(value)
            elif isinstance(value, time):
                return value

            return value

        rows: list[list[_CellValue]] = sheet.to_python(
            skip_empty_area=False, nrows=file_rows_needed
        )
        data = [[_convert_cell(cell) for cell in row] for row in rows]

        return data

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\models\target_python.py ===
import sys
from typing import List, Optional, Set, Tuple

from pip._vendor.packaging.tags import Tag

from pip._internal.utils.compatibility_tags import get_supported, version_info_to_nodot
from pip._internal.utils.misc import normalize_version_info


class TargetPython:
    """
    Encapsulates the properties of a Python interpreter one is targeting
    for a package install, download, etc.
    """

    __slots__ = [
        "_given_py_version_info",
        "abis",
        "implementation",
        "platforms",
        "py_version",
        "py_version_info",
        "_valid_tags",
        "_valid_tags_set",
    ]

    def __init__(
        self,
        platforms: Optional[List[str]] = None,
        py_version_info: Optional[Tuple[int, ...]] = None,
        abis: Optional[List[str]] = None,
        implementation: Optional[str] = None,
    ) -> None:
        """
        :param platforms: A list of strings or None. If None, searches for
            packages that are supported by the current system. Otherwise, will
            find packages that can be built on the platforms passed in. These
            packages will only be downloaded for distribution: they will
            not be built locally.
        :param py_version_info: An optional tuple of ints representing the
            Python version information to use (e.g. `sys.version_info[:3]`).
            This can have length 1, 2, or 3 when provided.
        :param abis: A list of strings or None. This is passed to
            compatibility_tags.py's get_supported() function as is.
        :param implementation: A string or None. This is passed to
            compatibility_tags.py's get_supported() function as is.
        """
        # Store the given py_version_info for when we call get_supported().
        self._given_py_version_info = py_version_info

        if py_version_info is None:
            py_version_info = sys.version_info[:3]
        else:
            py_version_info = normalize_version_info(py_version_info)

        py_version = ".".join(map(str, py_version_info[:2]))

        self.abis = abis
        self.implementation = implementation
        self.platforms = platforms
        self.py_version = py_version
        self.py_version_info = py_version_info

        # This is used to cache the return value of get_(un)sorted_tags.
        self._valid_tags: Optional[List[Tag]] = None
        self._valid_tags_set: Optional[Set[Tag]] = None

    def format_given(self) -> str:
        """
        Format the given, non-None attributes for display.
        """
        display_version = None
        if self._given_py_version_info is not None:
            display_version = ".".join(
                str(part) for part in self._given_py_version_info
            )

        key_values = [
            ("platforms", self.platforms),
            ("version_info", display_version),
            ("abis", self.abis),
            ("implementation", self.implementation),
        ]
        return " ".join(
            f"{key}={value!r}" for key, value in key_values if value is not None
        )

    def get_sorted_tags(self) -> List[Tag]:
        """
        Return the supported PEP 425 tags to check wheel candidates against.

        The tags are returned in order of preference (most preferred first).
        """
        if self._valid_tags is None:
            # Pass versions=None if no py_version_info was given since
            # versions=None uses special default logic.
            py_version_info = self._given_py_version_info
            if py_version_info is None:
                version = None
            else:
                version = version_info_to_nodot(py_version_info)

            tags = get_supported(
                version=version,
                platforms=self.platforms,
                abis=self.abis,
                impl=self.implementation,
            )
            self._valid_tags = tags

        return self._valid_tags

    def get_unsorted_tags(self) -> Set[Tag]:
        """Exactly the same as get_sorted_tags, but returns a set.

        This is important for performance.
        """
        if self._valid_tags_set is None:
            self._valid_tags_set = set(self.get_sorted_tags())

        return self._valid_tags_set

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\vector\deloperator.py ===
from sympy.core import Basic
from sympy.vector.operators import gradient, divergence, curl


class Del(Basic):
    """
    Represents the vector differential operator, usually represented in
    mathematical expressions as the 'nabla' symbol.
    """

    def __new__(cls):
        obj = super().__new__(cls)
        obj._name = "delop"
        return obj

    def gradient(self, scalar_field, doit=False):
        """
        Returns the gradient of the given scalar field, as a
        Vector instance.

        Parameters
        ==========

        scalar_field : SymPy expression
            The scalar field to calculate the gradient of.

        doit : bool
            If True, the result is returned after calling .doit() on
            each component. Else, the returned expression contains
            Derivative instances

        Examples
        ========

        >>> from sympy.vector import CoordSys3D, Del
        >>> C = CoordSys3D('C')
        >>> delop = Del()
        >>> delop.gradient(9)
        0
        >>> delop(C.x*C.y*C.z).doit()
        C.y*C.z*C.i + C.x*C.z*C.j + C.x*C.y*C.k

        """

        return gradient(scalar_field, doit=doit)

    __call__ = gradient
    __call__.__doc__ = gradient.__doc__

    def dot(self, vect, doit=False):
        """
        Represents the dot product between this operator and a given
        vector - equal to the divergence of the vector field.

        Parameters
        ==========

        vect : Vector
            The vector whose divergence is to be calculated.

        doit : bool
            If True, the result is returned after calling .doit() on
            each component. Else, the returned expression contains
            Derivative instances

        Examples
        ========

        >>> from sympy.vector import CoordSys3D, Del
        >>> delop = Del()
        >>> C = CoordSys3D('C')
        >>> delop.dot(C.x*C.i)
        Derivative(C.x, C.x)
        >>> v = C.x*C.y*C.z * (C.i + C.j + C.k)
        >>> (delop & v).doit()
        C.x*C.y + C.x*C.z + C.y*C.z

        """
        return divergence(vect, doit=doit)

    __and__ = dot
    __and__.__doc__ = dot.__doc__

    def cross(self, vect, doit=False):
        """
        Represents the cross product between this operator and a given
        vector - equal to the curl of the vector field.

        Parameters
        ==========

        vect : Vector
            The vector whose curl is to be calculated.

        doit : bool
            If True, the result is returned after calling .doit() on
            each component. Else, the returned expression contains
            Derivative instances

        Examples
        ========

        >>> from sympy.vector import CoordSys3D, Del
        >>> C = CoordSys3D('C')
        >>> delop = Del()
        >>> v = C.x*C.y*C.z * (C.i + C.j + C.k)
        >>> delop.cross(v, doit = True)
        (-C.x*C.y + C.x*C.z)*C.i + (C.x*C.y - C.y*C.z)*C.j +
            (-C.x*C.z + C.y*C.z)*C.k
        >>> (delop ^ C.i).doit()
        0

        """

        return curl(vect, doit=doit)

    __xor__ = cross
    __xor__.__doc__ = cross.__doc__

    def _sympystr(self, printer):
        return self._name

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_core\_dtype_ctypes.py ===
"""
Conversion from ctypes to dtype.

In an ideal world, we could achieve this through the PEP3118 buffer protocol,
something like::

    def dtype_from_ctypes_type(t):
        # needed to ensure that the shape of `t` is within memoryview.format
        class DummyStruct(ctypes.Structure):
            _fields_ = [('a', t)]

        # empty to avoid memory allocation
        ctype_0 = (DummyStruct * 0)()
        mv = memoryview(ctype_0)

        # convert the struct, and slice back out the field
        return _dtype_from_pep3118(mv.format)['a']

Unfortunately, this fails because:

* ctypes cannot handle length-0 arrays with PEP3118 (bpo-32782)
* PEP3118 cannot represent unions, but both numpy and ctypes can
* ctypes cannot handle big-endian structs with PEP3118 (bpo-32780)
"""

# We delay-import ctypes for distributions that do not include it.
# While this module is not used unless the user passes in ctypes
# members, it is eagerly imported from numpy/_core/__init__.py.
import numpy as np


def _from_ctypes_array(t):
    return np.dtype((dtype_from_ctypes_type(t._type_), (t._length_,)))


def _from_ctypes_structure(t):
    for item in t._fields_:
        if len(item) > 2:
            raise TypeError(
                "ctypes bitfields have no dtype equivalent")

    if hasattr(t, "_pack_"):
        import ctypes
        formats = []
        offsets = []
        names = []
        current_offset = 0
        for fname, ftyp in t._fields_:
            names.append(fname)
            formats.append(dtype_from_ctypes_type(ftyp))
            # Each type has a default offset, this is platform dependent
            # for some types.
            effective_pack = min(t._pack_, ctypes.alignment(ftyp))
            current_offset = (
                (current_offset + effective_pack - 1) // effective_pack
            ) * effective_pack
            offsets.append(current_offset)
            current_offset += ctypes.sizeof(ftyp)

        return np.dtype({
            "formats": formats,
            "offsets": offsets,
            "names": names,
            "itemsize": ctypes.sizeof(t)})
    else:
        fields = []
        for fname, ftyp in t._fields_:
            fields.append((fname, dtype_from_ctypes_type(ftyp)))

        # by default, ctypes structs are aligned
        return np.dtype(fields, align=True)


def _from_ctypes_scalar(t):
    """
    Return the dtype type with endianness included if it's the case
    """
    if getattr(t, '__ctype_be__', None) is t:
        return np.dtype('>' + t._type_)
    elif getattr(t, '__ctype_le__', None) is t:
        return np.dtype('<' + t._type_)
    else:
        return np.dtype(t._type_)


def _from_ctypes_union(t):
    import ctypes
    formats = []
    offsets = []
    names = []
    for fname, ftyp in t._fields_:
        names.append(fname)
        formats.append(dtype_from_ctypes_type(ftyp))
        offsets.append(0)  # Union fields are offset to 0

    return np.dtype({
        "formats": formats,
        "offsets": offsets,
        "names": names,
        "itemsize": ctypes.sizeof(t)})


def dtype_from_ctypes_type(t):
    """
    Construct a dtype object from a ctypes type
    """
    import _ctypes
    if issubclass(t, _ctypes.Array):
        return _from_ctypes_array(t)
    elif issubclass(t, _ctypes._Pointer):
        raise TypeError("ctypes pointers have no dtype equivalent")
    elif issubclass(t, _ctypes.Structure):
        return _from_ctypes_structure(t)
    elif issubclass(t, _ctypes.Union):
        return _from_ctypes_union(t)
    elif isinstance(getattr(t, '_type_', None), str):
        return _from_ctypes_scalar(t)
    else:
        raise NotImplementedError(
            f"Unknown ctypes type {t.__name__}")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\tools\ort_format_model\ort_flatbuffers_py\fbs\DeprecatedKernelCreateInfos.py ===
# automatically generated by the FlatBuffers compiler, do not modify

# namespace: fbs

import flatbuffers
from flatbuffers.compat import import_numpy
np = import_numpy()

# deprecated: no longer using kernel def hashes
class DeprecatedKernelCreateInfos(object):
    __slots__ = ['_tab']

    @classmethod
    def GetRootAs(cls, buf, offset=0):
        n = flatbuffers.encode.Get(flatbuffers.packer.uoffset, buf, offset)
        x = DeprecatedKernelCreateInfos()
        x.Init(buf, n + offset)
        return x

    @classmethod
    def GetRootAsDeprecatedKernelCreateInfos(cls, buf, offset=0):
        """This method is deprecated. Please switch to GetRootAs."""
        return cls.GetRootAs(buf, offset)
    @classmethod
    def DeprecatedKernelCreateInfosBufferHasIdentifier(cls, buf, offset, size_prefixed=False):
        return flatbuffers.util.BufferHasIdentifier(buf, offset, b"\x4F\x52\x54\x4D", size_prefixed=size_prefixed)

    # DeprecatedKernelCreateInfos
    def Init(self, buf, pos):
        self._tab = flatbuffers.table.Table(buf, pos)

    # DeprecatedKernelCreateInfos
    def NodeIndices(self, j):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(4))
        if o != 0:
            a = self._tab.Vector(o)
            return self._tab.Get(flatbuffers.number_types.Uint32Flags, a + flatbuffers.number_types.UOffsetTFlags.py_type(j * 4))
        return 0

    # DeprecatedKernelCreateInfos
    def NodeIndicesAsNumpy(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(4))
        if o != 0:
            return self._tab.GetVectorAsNumpy(flatbuffers.number_types.Uint32Flags, o)
        return 0

    # DeprecatedKernelCreateInfos
    def NodeIndicesLength(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(4))
        if o != 0:
            return self._tab.VectorLen(o)
        return 0

    # DeprecatedKernelCreateInfos
    def NodeIndicesIsNone(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(4))
        return o == 0

    # DeprecatedKernelCreateInfos
    def KernelDefHashes(self, j):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        if o != 0:
            a = self._tab.Vector(o)
            return self._tab.Get(flatbuffers.number_types.Uint64Flags, a + flatbuffers.number_types.UOffsetTFlags.py_type(j * 8))
        return 0

    # DeprecatedKernelCreateInfos
    def KernelDefHashesAsNumpy(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        if o != 0:
            return self._tab.GetVectorAsNumpy(flatbuffers.number_types.Uint64Flags, o)
        return 0

    # DeprecatedKernelCreateInfos
    def KernelDefHashesLength(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        if o != 0:
            return self._tab.VectorLen(o)
        return 0

    # DeprecatedKernelCreateInfos
    def KernelDefHashesIsNone(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        return o == 0

def DeprecatedKernelCreateInfosStart(builder):
    builder.StartObject(2)

def Start(builder):
    DeprecatedKernelCreateInfosStart(builder)

def DeprecatedKernelCreateInfosAddNodeIndices(builder, nodeIndices):
    builder.PrependUOffsetTRelativeSlot(0, flatbuffers.number_types.UOffsetTFlags.py_type(nodeIndices), 0)

def AddNodeIndices(builder, nodeIndices):
    DeprecatedKernelCreateInfosAddNodeIndices(builder, nodeIndices)

def DeprecatedKernelCreateInfosStartNodeIndicesVector(builder, numElems):
    return builder.StartVector(4, numElems, 4)

def StartNodeIndicesVector(builder, numElems: int) -> int:
    return DeprecatedKernelCreateInfosStartNodeIndicesVector(builder, numElems)

def DeprecatedKernelCreateInfosAddKernelDefHashes(builder, kernelDefHashes):
    builder.PrependUOffsetTRelativeSlot(1, flatbuffers.number_types.UOffsetTFlags.py_type(kernelDefHashes), 0)

def AddKernelDefHashes(builder, kernelDefHashes):
    DeprecatedKernelCreateInfosAddKernelDefHashes(builder, kernelDefHashes)

def DeprecatedKernelCreateInfosStartKernelDefHashesVector(builder, numElems):
    return builder.StartVector(8, numElems, 8)

def StartKernelDefHashesVector(builder, numElems: int) -> int:
    return DeprecatedKernelCreateInfosStartKernelDefHashesVector(builder, numElems)

def DeprecatedKernelCreateInfosEnd(builder):
    return builder.EndObject()

def End(builder):
    return DeprecatedKernelCreateInfosEnd(builder)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\worksheet\protection.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.descriptors import (
    Bool,
    String,
    Alias,
    Integer,
)
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors.excel import (
    Base64Binary,
)
from openpyxl.utils.protection import hash_password


class _Protected:
    _password = None

    def set_password(self, value='', already_hashed=False):
        """Set a password on this sheet."""
        if not already_hashed:
            value = hash_password(value)
        self._password = value

    @property
    def password(self):
        """Return the password value, regardless of hash."""
        return self._password

    @password.setter
    def password(self, value):
        """Set a password directly, forcing a hash step."""
        self.set_password(value)


class SheetProtection(Serialisable, _Protected):
    """
    Information about protection of various aspects of a sheet. True values
    mean that protection for the object or action is active This is the
    **default** when protection is active, ie. users cannot do something
    """

    tagname = "sheetProtection"

    sheet = Bool()
    enabled = Alias('sheet')
    objects = Bool()
    scenarios = Bool()
    formatCells = Bool()
    formatColumns = Bool()
    formatRows = Bool()
    insertColumns = Bool()
    insertRows = Bool()
    insertHyperlinks = Bool()
    deleteColumns = Bool()
    deleteRows = Bool()
    selectLockedCells = Bool()
    selectUnlockedCells = Bool()
    sort = Bool()
    autoFilter = Bool()
    pivotTables = Bool()
    saltValue = Base64Binary(allow_none=True)
    spinCount = Integer(allow_none=True)
    algorithmName = String(allow_none=True)
    hashValue = Base64Binary(allow_none=True)


    __attrs__ = ('selectLockedCells', 'selectUnlockedCells', 'algorithmName',
              'sheet', 'objects', 'insertRows', 'insertHyperlinks', 'autoFilter',
              'scenarios', 'formatColumns', 'deleteColumns', 'insertColumns',
              'pivotTables', 'deleteRows', 'formatCells', 'saltValue', 'formatRows',
              'sort', 'spinCount', 'password', 'hashValue')


    def __init__(self, sheet=False, objects=False, scenarios=False,
                 formatCells=True, formatRows=True, formatColumns=True,
                 insertColumns=True, insertRows=True, insertHyperlinks=True,
                 deleteColumns=True, deleteRows=True, selectLockedCells=False,
                 selectUnlockedCells=False, sort=True, autoFilter=True, pivotTables=True,
                 password=None, algorithmName=None, saltValue=None, spinCount=None, hashValue=None):
        self.sheet = sheet
        self.objects = objects
        self.scenarios = scenarios
        self.formatCells = formatCells
        self.formatColumns = formatColumns
        self.formatRows = formatRows
        self.insertColumns = insertColumns
        self.insertRows = insertRows
        self.insertHyperlinks = insertHyperlinks
        self.deleteColumns = deleteColumns
        self.deleteRows = deleteRows
        self.selectLockedCells = selectLockedCells
        self.selectUnlockedCells = selectUnlockedCells
        self.sort = sort
        self.autoFilter = autoFilter
        self.pivotTables = pivotTables
        if password is not None:
            self.password = password
        self.algorithmName = algorithmName
        self.saltValue = saltValue
        self.spinCount = spinCount
        self.hashValue = hashValue


    def set_password(self, value='', already_hashed=False):
        super().set_password(value, already_hashed)
        self.enable()


    def enable(self):
        self.sheet = True


    def disable(self):
        self.sheet = False


    def  __bool__(self):
        return self.sheet


# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\indexes\frozen.py ===
"""
frozen (immutable) data structures to support MultiIndexing

These are used for:

- .names (FrozenList)

"""
from __future__ import annotations

from typing import (
    TYPE_CHECKING,
    NoReturn,
)

from pandas.core.base import PandasObject

from pandas.io.formats.printing import pprint_thing

if TYPE_CHECKING:
    from pandas._typing import Self


class FrozenList(PandasObject, list):
    """
    Container that doesn't allow setting item *but*
    because it's technically hashable, will be used
    for lookups, appropriately, etc.
    """

    # Side note: This has to be of type list. Otherwise,
    #            it messes up PyTables type checks.

    def union(self, other) -> FrozenList:
        """
        Returns a FrozenList with other concatenated to the end of self.

        Parameters
        ----------
        other : array-like
            The array-like whose elements we are concatenating.

        Returns
        -------
        FrozenList
            The collection difference between self and other.
        """
        if isinstance(other, tuple):
            other = list(other)
        return type(self)(super().__add__(other))

    def difference(self, other) -> FrozenList:
        """
        Returns a FrozenList with elements from other removed from self.

        Parameters
        ----------
        other : array-like
            The array-like whose elements we are removing self.

        Returns
        -------
        FrozenList
            The collection difference between self and other.
        """
        other = set(other)
        temp = [x for x in self if x not in other]
        return type(self)(temp)

    # TODO: Consider deprecating these in favor of `union` (xref gh-15506)
    # error: Incompatible types in assignment (expression has type
    # "Callable[[FrozenList, Any], FrozenList]", base class "list" defined the
    # type as overloaded function)
    __add__ = __iadd__ = union  # type: ignore[assignment]

    def __getitem__(self, n):
        if isinstance(n, slice):
            return type(self)(super().__getitem__(n))
        return super().__getitem__(n)

    def __radd__(self, other) -> Self:
        if isinstance(other, tuple):
            other = list(other)
        return type(self)(other + list(self))

    def __eq__(self, other: object) -> bool:
        if isinstance(other, (tuple, FrozenList)):
            other = list(other)
        return super().__eq__(other)

    __req__ = __eq__

    def __mul__(self, other) -> Self:
        return type(self)(super().__mul__(other))

    __imul__ = __mul__

    def __reduce__(self):
        return type(self), (list(self),)

    # error: Signature of "__hash__" incompatible with supertype "list"
    def __hash__(self) -> int:  # type: ignore[override]
        return hash(tuple(self))

    def _disabled(self, *args, **kwargs) -> NoReturn:
        """
        This method will not function because object is immutable.
        """
        raise TypeError(f"'{type(self).__name__}' does not support mutable operations.")

    def __str__(self) -> str:
        return pprint_thing(self, quote_strings=True, escape_chars=("\t", "\r", "\n"))

    def __repr__(self) -> str:
        return f"{type(self).__name__}({str(self)})"

    __setitem__ = __setslice__ = _disabled  # type: ignore[assignment]
    __delitem__ = __delslice__ = _disabled
    pop = append = extend = _disabled
    remove = sort = insert = _disabled  # type: ignore[assignment]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\util\topological.py ===
# util/topological.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

"""Topological sorting algorithms."""

from __future__ import annotations

from typing import Any
from typing import Collection
from typing import DefaultDict
from typing import Iterable
from typing import Iterator
from typing import Sequence
from typing import Set
from typing import Tuple
from typing import TypeVar

from .. import util
from ..exc import CircularDependencyError

_T = TypeVar("_T", bound=Any)

__all__ = ["sort", "sort_as_subsets", "find_cycles"]


def sort_as_subsets(
    tuples: Collection[Tuple[_T, _T]], allitems: Collection[_T]
) -> Iterator[Sequence[_T]]:
    edges: DefaultDict[_T, Set[_T]] = util.defaultdict(set)
    for parent, child in tuples:
        edges[child].add(parent)

    todo = list(allitems)
    todo_set = set(allitems)

    while todo_set:
        output = []
        for node in todo:
            if todo_set.isdisjoint(edges[node]):
                output.append(node)

        if not output:
            raise CircularDependencyError(
                "Circular dependency detected.",
                find_cycles(tuples, allitems),
                _gen_edges(edges),
            )

        todo_set.difference_update(output)
        todo = [t for t in todo if t in todo_set]
        yield output


def sort(
    tuples: Collection[Tuple[_T, _T]],
    allitems: Collection[_T],
    deterministic_order: bool = True,
) -> Iterator[_T]:
    """sort the given list of items by dependency.

    'tuples' is a list of tuples representing a partial ordering.

    deterministic_order is no longer used, the order is now always
    deterministic given the order of "allitems".    the flag is there
    for backwards compatibility with Alembic.

    """

    for set_ in sort_as_subsets(tuples, allitems):
        yield from set_


def find_cycles(
    tuples: Iterable[Tuple[_T, _T]], allitems: Iterable[_T]
) -> Set[_T]:
    # adapted from:
    # https://neopythonic.blogspot.com/2009/01/detecting-cycles-in-directed-graph.html

    edges: DefaultDict[_T, Set[_T]] = util.defaultdict(set)
    for parent, child in tuples:
        edges[parent].add(child)
    nodes_to_test = set(edges)

    output = set()

    # we'd like to find all nodes that are
    # involved in cycles, so we do the full
    # pass through the whole thing for each
    # node in the original list.

    # we can go just through parent edge nodes.
    # if a node is only a child and never a parent,
    # by definition it can't be part of a cycle.  same
    # if it's not in the edges at all.
    for node in nodes_to_test:
        stack = [node]
        todo = nodes_to_test.difference(stack)
        while stack:
            top = stack[-1]
            for node in edges[top]:
                if node in stack:
                    cyc = stack[stack.index(node) :]
                    todo.difference_update(cyc)
                    output.update(cyc)

                if node in todo:
                    stack.append(node)
                    todo.remove(node)
                    break
            else:
                stack.pop()
    return output


def _gen_edges(edges: DefaultDict[_T, Set[_T]]) -> Set[Tuple[_T, _T]]:
    return {(right, left) for left in edges for right in edges[left]}

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\core\backend.py ===
import os
USE_SYMENGINE = os.getenv('USE_SYMENGINE', '0')
USE_SYMENGINE = USE_SYMENGINE.lower() in ('1', 't', 'true')  # type: ignore

if USE_SYMENGINE:
    from symengine import (Symbol, Integer, sympify as sympify_symengine, S,
        SympifyError, exp, log, gamma, sqrt, I, E, pi, Matrix,
        sin, cos, tan, cot, csc, sec, asin, acos, atan, acot, acsc, asec,
        sinh, cosh, tanh, coth, asinh, acosh, atanh, acoth,
        lambdify, symarray, diff, zeros, eye, diag, ones,
        expand, Function, symbols, var, Add, Mul, Derivative,
        ImmutableMatrix, MatrixBase, Rational, Basic)
    from symengine.lib.symengine_wrapper import gcd as igcd
    from symengine import AppliedUndef

    def sympify(a, *, strict=False):
        """
        Notes
        =====

        SymEngine's ``sympify`` does not accept keyword arguments and is
        therefore not compatible with SymPy's ``sympify`` with ``strict=True``
        (which ensures that only the types for which an explicit conversion has
        been defined are converted). This wrapper adds an additional parameter
        ``strict`` (with default ``False``) that will raise a ``SympifyError``
        if ``strict=True`` and the argument passed to the parameter ``a`` is a
        string.

        See Also
        ========

        sympify: Converts an arbitrary expression to a type that can be used
            inside SymPy.

        """
        # The parameter ``a`` is used for this function to keep compatibility
        # with the SymEngine docstring.
        if strict and isinstance(a, str):
            raise SympifyError(a)
        return sympify_symengine(a)

    # Keep the SymEngine docstring and append the additional "Notes" and "See
    # Also" sections. Replacement of spaces is required to correctly format the
    # indentation of the combined docstring.
    sympify.__doc__ = (
        sympify_symengine.__doc__
        + sympify.__doc__.replace('        ', '    ')  # type: ignore
    )
else:
    from sympy.core.add import Add
    from sympy.core.basic import Basic
    from sympy.core.function import (diff, Function, AppliedUndef,
        expand, Derivative)
    from sympy.core.mul import Mul
    from sympy.core.intfunc import igcd
    from sympy.core.numbers import pi, I, Integer, Rational, E
    from sympy.core.singleton import S
    from sympy.core.symbol import Symbol, var, symbols
    from sympy.core.sympify import SympifyError, sympify
    from sympy.functions.elementary.exponential import log, exp
    from sympy.functions.elementary.hyperbolic import (coth, sinh,
        acosh, acoth, tanh, asinh, atanh, cosh)
    from sympy.functions.elementary.miscellaneous import sqrt
    from sympy.functions.elementary.trigonometric import (csc,
        asec, cos, atan, sec, acot, asin, tan, sin, cot, acsc, acos)
    from sympy.functions.special.gamma_functions import gamma
    from sympy.matrices.dense import (eye, zeros, diag, Matrix,
        ones, symarray)
    from sympy.matrices.immutable import ImmutableMatrix
    from sympy.matrices.matrixbase import MatrixBase
    from sympy.utilities.lambdify import lambdify


#
# XXX: Handling of immutable and mutable matrices in SymEngine is inconsistent
# with SymPy's matrix classes in at least SymEngine version 0.7.0. Until that
# is fixed the function below is needed for consistent behaviour when
# attempting to simplify a matrix.
#
# Expected behaviour of a SymPy mutable/immutable matrix .simplify() method:
#
#   Matrix.simplify() : works in place, returns None
#   ImmutableMatrix.simplify() : returns a simplified copy
#
# In SymEngine both mutable and immutable matrices simplify in place and return
# None. This is inconsistent with the matrix being "immutable" and also the
# returned None leads to problems in the mechanics module.
#
# The simplify function should not be used because simplify(M) sympifies the
# matrix M and the SymEngine matrices all sympify to SymPy matrices. If we want
# to work with SymEngine matrices then we need to use their .simplify() method
# but that method does not work correctly with immutable matrices.
#
# The _simplify_matrix function can be removed when the SymEngine bug is fixed.
# Since this should be a temporary problem we do not make this function part of
# the public API.
#
#   SymEngine issue: https://github.com/symengine/symengine.py/issues/363
#

def _simplify_matrix(M):
    """Return a simplified copy of the matrix M"""
    if not isinstance(M, (Matrix, ImmutableMatrix)):
        raise TypeError("The matrix M must be an instance of Matrix or ImmutableMatrix")
    Mnew = M.as_mutable() # makes a copy if mutable
    Mnew.simplify()
    if isinstance(M, ImmutableMatrix):
        Mnew = Mnew.as_immutable()
    return Mnew


__all__ = [
    'Symbol', 'Integer', 'sympify', 'S', 'SympifyError', 'exp', 'log',
    'gamma', 'sqrt', 'I', 'E', 'pi', 'Matrix', 'sin', 'cos', 'tan', 'cot',
    'csc', 'sec', 'asin', 'acos', 'atan', 'acot', 'acsc', 'asec', 'sinh',
    'cosh', 'tanh', 'coth', 'asinh', 'acosh', 'atanh', 'acoth', 'lambdify',
    'symarray', 'diff', 'zeros', 'eye', 'diag', 'ones', 'expand', 'Function',
    'symbols', 'var', 'Add', 'Mul', 'Derivative', 'ImmutableMatrix',
    'MatrixBase', 'Rational', 'Basic', 'igcd', 'AppliedUndef',
]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_core\_type_aliases.py ===
"""
Due to compatibility, numpy has a very large number of different naming
conventions for the scalar types (those subclassing from `numpy.generic`).
This file produces a convoluted set of dictionaries mapping names to types,
and sometimes other mappings too.

.. data:: allTypes
    A dictionary of names to types that will be exposed as attributes through
    ``np._core.numerictypes.*``

.. data:: sctypeDict
    Similar to `allTypes`, but maps a broader set of aliases to their types.

.. data:: sctypes
    A dictionary keyed by a "type group" string, providing a list of types
    under that group.

"""

import numpy._core.multiarray as ma
from numpy._core.multiarray import dtype, typeinfo

######################################
# Building `sctypeDict` and `allTypes`
######################################

sctypeDict = {}
allTypes = {}
c_names_dict = {}

_abstract_type_names = {
    "generic", "integer", "inexact", "floating", "number",
    "flexible", "character", "complexfloating", "unsignedinteger",
    "signedinteger"
}

for _abstract_type_name in _abstract_type_names:
    allTypes[_abstract_type_name] = getattr(ma, _abstract_type_name)

for k, v in typeinfo.items():
    if k.startswith("NPY_") and v not in c_names_dict:
        c_names_dict[k[4:]] = v
    else:
        concrete_type = v.type
        allTypes[k] = concrete_type
        sctypeDict[k] = concrete_type

_aliases = {
    "double": "float64",
    "cdouble": "complex128",
    "single": "float32",
    "csingle": "complex64",
    "half": "float16",
    "bool_": "bool",
    # Default integer:
    "int_": "intp",
    "uint": "uintp",
}

for k, v in _aliases.items():
    sctypeDict[k] = allTypes[v]
    allTypes[k] = allTypes[v]

# extra aliases are added only to `sctypeDict`
# to support dtype name access, such as`np.dtype("float")`
_extra_aliases = {
    "float": "float64",
    "complex": "complex128",
    "object": "object_",
    "bytes": "bytes_",
    "a": "bytes_",
    "int": "int_",
    "str": "str_",
    "unicode": "str_",
}

for k, v in _extra_aliases.items():
    sctypeDict[k] = allTypes[v]

# include extended precision sized aliases
for is_complex, full_name in [(False, "longdouble"), (True, "clongdouble")]:
    longdouble_type: type = allTypes[full_name]

    bits: int = dtype(longdouble_type).itemsize * 8
    base_name: str = "complex" if is_complex else "float"
    extended_prec_name: str = f"{base_name}{bits}"
    if extended_prec_name not in allTypes:
        sctypeDict[extended_prec_name] = longdouble_type
        allTypes[extended_prec_name] = longdouble_type


####################
# Building `sctypes`
####################

sctypes = {"int": set(), "uint": set(), "float": set(),
           "complex": set(), "others": set()}

for type_info in typeinfo.values():
    if type_info.kind in ["M", "m"]:  # exclude timedelta and datetime
        continue

    concrete_type = type_info.type

    # find proper group for each concrete type
    for type_group, abstract_type in [
        ("int", ma.signedinteger), ("uint", ma.unsignedinteger),
        ("float", ma.floating), ("complex", ma.complexfloating),
        ("others", ma.generic)
    ]:
        if issubclass(concrete_type, abstract_type):
            sctypes[type_group].add(concrete_type)
            break

# sort sctype groups by bitsize
for sctype_key in sctypes.keys():
    sctype_list = list(sctypes[sctype_key])
    sctype_list.sort(key=lambda x: dtype(x).itemsize)
    sctypes[sctype_key] = sctype_list

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\quantization\operators\activation.py ===
import onnx

from ..quant_utils import TENSOR_NAME_QUANT_SUFFIX, QuantizedValue, QuantizedValueType, attribute_to_kwarg, ms_domain
from .base_operator import QuantOperatorBase
from .qdq_base_operator import QDQOperatorBase


class QLinearActivation(QuantOperatorBase):
    def __init__(self, onnx_quantizer, onnx_node):
        super().__init__(onnx_quantizer, onnx_node)

    def QuantizeClipRelu(self):  # noqa: N802
        node = self.node
        assert node.op_type == "Relu" or node.op_type == "Clip"

        # When mode is QLinearOps, the output quantization params are calculated based on outputs from
        # activation nodes, therefore these nodes can be removed from the graph if they follow a quantized op.
        # If input to this node is not quantized then keep this node
        # If activation is symmetric, not quantize the op and simply return
        if node.input[0] not in self.quantizer.quantized_value_map or self.quantizer.is_activation_symmetric:
            return super().quantize()

        quantized_value = self.quantizer.quantized_value_map[node.input[0]]
        self.quantizer.quantized_value_map[node.output[0]] = quantized_value

    def quantize(self):
        node = self.node
        if node.op_type == "Relu" or node.op_type == "Clip":
            self.QuantizeClipRelu()
            return

        nnapi_sigmoid_option = "extra.Sigmoid.nnapi"
        sigmoid_nnapi_mode = (
            node.op_type == "Sigmoid"
            and nnapi_sigmoid_option in self.quantizer.extra_options
            and self.quantizer.extra_options[nnapi_sigmoid_option]
        )
        use_scale = 1 / 256.0 if sigmoid_nnapi_mode else None
        use_zeropoint = 0 if sigmoid_nnapi_mode else None

        # No assert on op_type as it is controlled by registry
        # only try to quantize when given quantization parameters for it
        (
            data_found,
            output_scale_name,
            output_zp_name,
            _,
            _,
        ) = self.quantizer._get_quantization_params(node.output[0], use_scale, use_zeropoint)
        (
            quantized_input_names,
            zero_point_names,
            scale_names,
            nodes,
        ) = self.quantizer.quantize_activation(node, [0])
        if not data_found or quantized_input_names is None:
            return super().quantize()

        qlinear_activation_output = node.output[0] + TENSOR_NAME_QUANT_SUFFIX
        qlinear_activation_name = ""
        if node.name:
            qlinear_activation_name = node.name + "_quant"
        kwargs = {}
        for attribute in node.attribute:
            kwargs.update(attribute_to_kwarg(attribute))
        kwargs["domain"] = ms_domain

        qlinear_activation_inputs = [
            quantized_input_names[0],
            scale_names[0],
            zero_point_names[0],
            output_scale_name,
            output_zp_name,
        ]

        qlinear_activation_node = onnx.helper.make_node(
            "QLinear" + node.op_type,
            qlinear_activation_inputs,
            [qlinear_activation_output],
            qlinear_activation_name,
            **kwargs,
        )

        # Create an entry for this quantized value
        q_output = QuantizedValue(
            node.output[0],
            qlinear_activation_output,
            output_scale_name,
            output_zp_name,
            QuantizedValueType.Input,
        )
        self.quantizer.quantized_value_map[node.output[0]] = q_output

        nodes.append(qlinear_activation_node)
        self.quantizer.new_nodes += nodes


class QDQRemovableActivation(QDQOperatorBase):
    def __init__(self, onnx_quantizer, onnx_node):
        super().__init__(onnx_quantizer, onnx_node)

    def quantize(self):
        node = self.node

        # If input to this node is not quantized then keep this node
        if not self.quantizer.is_tensor_quantized(node.input[0]):
            return

        if (
            not self.quantizer.is_activation_symmetric
            and not self.quantizer.qdq_keep_removable_activations
            and self.quantizer.try_replacing_upstream_output(node.input[0], node.output[0])
        ):
            self.quantizer.remove_node(self.node)
        else:
            self.quantizer.quantize_activation_tensor(node.input[0])

        if not self.disable_qdq_for_node_output:
            self.quantizer.quantize_activation_tensor(node.output[0])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\chart\surface_chart.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Typed,
    Integer,
    Bool,
    Alias,
    Sequence,
)
from openpyxl.descriptors.excel import ExtensionList
from openpyxl.descriptors.nested import (
    NestedInteger,
    NestedBool,
)

from ._chart import ChartBase
from ._3d import _3DBase
from .axis import TextAxis, NumericAxis, SeriesAxis
from .shapes import GraphicalProperties
from .series import Series


class BandFormat(Serialisable):

    tagname = "bandFmt"

    idx = NestedInteger()
    spPr = Typed(expected_type=GraphicalProperties, allow_none=True)
    graphicalProperties = Alias("spPr")

    __elements__ = ('idx', 'spPr')

    def __init__(self,
                 idx=0,
                 spPr=None,
                ):
        self.idx = idx
        self.spPr = spPr


class BandFormatList(Serialisable):

    tagname = "bandFmts"

    bandFmt = Sequence(expected_type=BandFormat, allow_none=True)

    __elements__ = ('bandFmt',)

    def __init__(self,
                 bandFmt=(),
                ):
        self.bandFmt = bandFmt


class _SurfaceChartBase(ChartBase):

    wireframe = NestedBool(allow_none=True)
    ser = Sequence(expected_type=Series, allow_none=True)
    bandFmts = Typed(expected_type=BandFormatList, allow_none=True)

    _series_type = "surface"

    __elements__ = ('wireframe', 'ser', 'bandFmts')

    def __init__(self,
                 wireframe=None,
                 ser=(),
                 bandFmts=None,
                 **kw
                ):
        self.wireframe = wireframe
        self.ser = ser
        self.bandFmts = bandFmts
        super().__init__(**kw)


class SurfaceChart3D(_SurfaceChartBase, _3DBase):

    tagname = "surface3DChart"

    wireframe = _SurfaceChartBase.wireframe
    ser = _SurfaceChartBase.ser
    bandFmts = _SurfaceChartBase.bandFmts

    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    x_axis = Typed(expected_type=TextAxis)
    y_axis = Typed(expected_type=NumericAxis)
    z_axis = Typed(expected_type=SeriesAxis)

    __elements__ = _SurfaceChartBase.__elements__ + ('axId',)

    def __init__(self, **kw):
        self.x_axis = TextAxis()
        self.y_axis = NumericAxis()
        self.z_axis = SeriesAxis()
        super(SurfaceChart3D, self).__init__(**kw)


class SurfaceChart(SurfaceChart3D):

    tagname = "surfaceChart"

    wireframe = _SurfaceChartBase.wireframe
    ser = _SurfaceChartBase.ser
    bandFmts = _SurfaceChartBase.bandFmts

    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = SurfaceChart3D.__elements__

    def __init__(self, **kw):
        super().__init__(**kw)
        self.y_axis.delete = True
        self.view3D.x_rotation = 90
        self.view3D.y_rotation = 0
        self.view3D.perspective = False
        self.view3D.right_angle_axes = False

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\cachecontrol\filewrapper.py ===
# SPDX-FileCopyrightText: 2015 Eric Larson
#
# SPDX-License-Identifier: Apache-2.0
from __future__ import annotations

import mmap
from tempfile import NamedTemporaryFile
from typing import TYPE_CHECKING, Any, Callable

if TYPE_CHECKING:
    from http.client import HTTPResponse


class CallbackFileWrapper:
    """
    Small wrapper around a fp object which will tee everything read into a
    buffer, and when that file is closed it will execute a callback with the
    contents of that buffer.

    All attributes are proxied to the underlying file object.

    This class uses members with a double underscore (__) leading prefix so as
    not to accidentally shadow an attribute.

    The data is stored in a temporary file until it is all available.  As long
    as the temporary files directory is disk-based (sometimes it's a
    memory-backed-``tmpfs`` on Linux), data will be unloaded to disk if memory
    pressure is high.  For small files the disk usually won't be used at all,
    it'll all be in the filesystem memory cache, so there should be no
    performance impact.
    """

    def __init__(
        self, fp: HTTPResponse, callback: Callable[[bytes], None] | None
    ) -> None:
        self.__buf = NamedTemporaryFile("rb+", delete=True)
        self.__fp = fp
        self.__callback = callback

    def __getattr__(self, name: str) -> Any:
        # The vagaries of garbage collection means that self.__fp is
        # not always set.  By using __getattribute__ and the private
        # name[0] allows looking up the attribute value and raising an
        # AttributeError when it doesn't exist. This stop things from
        # infinitely recursing calls to getattr in the case where
        # self.__fp hasn't been set.
        #
        # [0] https://docs.python.org/2/reference/expressions.html#atom-identifiers
        fp = self.__getattribute__("_CallbackFileWrapper__fp")
        return getattr(fp, name)

    def __is_fp_closed(self) -> bool:
        try:
            return self.__fp.fp is None

        except AttributeError:
            pass

        try:
            closed: bool = self.__fp.closed
            return closed

        except AttributeError:
            pass

        # We just don't cache it then.
        # TODO: Add some logging here...
        return False

    def _close(self) -> None:
        if self.__callback:
            if self.__buf.tell() == 0:
                # Empty file:
                result = b""
            else:
                # Return the data without actually loading it into memory,
                # relying on Python's buffer API and mmap(). mmap() just gives
                # a view directly into the filesystem's memory cache, so it
                # doesn't result in duplicate memory use.
                self.__buf.seek(0, 0)
                result = memoryview(
                    mmap.mmap(self.__buf.fileno(), 0, access=mmap.ACCESS_READ)
                )
            self.__callback(result)

        # We assign this to None here, because otherwise we can get into
        # really tricky problems where the CPython interpreter dead locks
        # because the callback is holding a reference to something which
        # has a __del__ method. Setting this to None breaks the cycle
        # and allows the garbage collector to do it's thing normally.
        self.__callback = None

        # Closing the temporary file releases memory and frees disk space.
        # Important when caching big files.
        self.__buf.close()

    def read(self, amt: int | None = None) -> bytes:
        data: bytes = self.__fp.read(amt)
        if data:
            # We may be dealing with b'', a sign that things are over:
            # it's passed e.g. after we've already closed self.__buf.
            self.__buf.write(data)
        if self.__is_fp_closed():
            self._close()

        return data

    def _safe_read(self, amt: int) -> bytes:
        data: bytes = self.__fp._safe_read(amt)  # type: ignore[attr-defined]
        if amt == 2 and data == b"\r\n":
            # urllib executes this read to toss the CRLF at the end
            # of the chunk.
            return data

        self.__buf.write(data)
        if self.__is_fp_closed():
            self._close()

        return data

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\sets\powerset.py ===
from sympy.core.decorators import _sympifyit
from sympy.core.parameters import global_parameters
from sympy.core.logic import fuzzy_bool
from sympy.core.singleton import S
from sympy.core.sympify import _sympify

from .sets import Set, FiniteSet, SetKind


class PowerSet(Set):
    r"""A symbolic object representing a power set.

    Parameters
    ==========

    arg : Set
        The set to take power of.

    evaluate : bool
        The flag to control evaluation.

        If the evaluation is disabled for finite sets, it can take
        advantage of using subset test as a membership test.

    Notes
    =====

    Power set `\mathcal{P}(S)` is defined as a set containing all the
    subsets of `S`.

    If the set `S` is a finite set, its power set would have
    `2^{\left| S \right|}` elements, where `\left| S \right|` denotes
    the cardinality of `S`.

    Examples
    ========

    >>> from sympy import PowerSet, S, FiniteSet

    A power set of a finite set:

    >>> PowerSet(FiniteSet(1, 2, 3))
    PowerSet({1, 2, 3})

    A power set of an empty set:

    >>> PowerSet(S.EmptySet)
    PowerSet(EmptySet)
    >>> PowerSet(PowerSet(S.EmptySet))
    PowerSet(PowerSet(EmptySet))

    A power set of an infinite set:

    >>> PowerSet(S.Reals)
    PowerSet(Reals)

    Evaluating the power set of a finite set to its explicit form:

    >>> PowerSet(FiniteSet(1, 2, 3)).rewrite(FiniteSet)
    FiniteSet(EmptySet, {1}, {2}, {3}, {1, 2}, {1, 3}, {2, 3}, {1, 2, 3})

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Power_set

    .. [2] https://en.wikipedia.org/wiki/Axiom_of_power_set
    """
    def __new__(cls, arg, evaluate=None):
        if evaluate is None:
            evaluate=global_parameters.evaluate

        arg = _sympify(arg)

        if not isinstance(arg, Set):
            raise ValueError('{} must be a set.'.format(arg))

        return super().__new__(cls, arg)

    @property
    def arg(self):
        return self.args[0]

    def _eval_rewrite_as_FiniteSet(self, *args, **kwargs):
        arg = self.arg
        if arg.is_FiniteSet:
            return arg.powerset()
        return None

    @_sympifyit('other', NotImplemented)
    def _contains(self, other):
        if not isinstance(other, Set):
            return None

        return fuzzy_bool(self.arg.is_superset(other))

    def _eval_is_subset(self, other):
        if isinstance(other, PowerSet):
            return self.arg.is_subset(other.arg)

    def __len__(self):
        return 2 ** len(self.arg)

    def __iter__(self):
        found = [S.EmptySet]
        yield S.EmptySet

        for x in self.arg:
            temp = []
            x = FiniteSet(x)
            for y in found:
                new = x + y
                yield new
                temp.append(new)
            found.extend(temp)

    @property
    def kind(self):
        return SetKind(self.arg.kind)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\google\protobuf\internal\builder.py ===
# Protocol Buffers - Google's data interchange format
# Copyright 2008 Google Inc.  All rights reserved.
#
# Use of this source code is governed by a BSD-style
# license that can be found in the LICENSE file or at
# https://developers.google.com/open-source/licenses/bsd

"""Builds descriptors, message classes and services for generated _pb2.py.

This file is only called in python generated _pb2.py files. It builds
descriptors, message classes and services that users can directly use
in generated code.
"""

__author__ = 'jieluo@google.com (Jie Luo)'

from google.protobuf.internal import enum_type_wrapper
from google.protobuf.internal import python_message
from google.protobuf import message as _message
from google.protobuf import reflection as _reflection
from google.protobuf import symbol_database as _symbol_database

_sym_db = _symbol_database.Default()


def BuildMessageAndEnumDescriptors(file_des, module):
  """Builds message and enum descriptors.

  Args:
    file_des: FileDescriptor of the .proto file
    module: Generated _pb2 module
  """

  def BuildNestedDescriptors(msg_des, prefix):
    for (name, nested_msg) in msg_des.nested_types_by_name.items():
      module_name = prefix + name.upper()
      module[module_name] = nested_msg
      BuildNestedDescriptors(nested_msg, module_name + '_')
    for enum_des in msg_des.enum_types:
      module[prefix + enum_des.name.upper()] = enum_des

  for (name, msg_des) in file_des.message_types_by_name.items():
    module_name = '_' + name.upper()
    module[module_name] = msg_des
    BuildNestedDescriptors(msg_des, module_name + '_')


def BuildTopDescriptorsAndMessages(file_des, module_name, module):
  """Builds top level descriptors and message classes.

  Args:
    file_des: FileDescriptor of the .proto file
    module_name: str, the name of generated _pb2 module
    module: Generated _pb2 module
  """

  def BuildMessage(msg_des, prefix):
    create_dict = {}
    for (name, nested_msg) in msg_des.nested_types_by_name.items():
      create_dict[name] = BuildMessage(nested_msg, prefix + msg_des.name + '.')
    create_dict['DESCRIPTOR'] = msg_des
    create_dict['__module__'] = module_name
    create_dict['__qualname__'] = prefix + msg_des.name
    message_class = _reflection.GeneratedProtocolMessageType(
        msg_des.name, (_message.Message,), create_dict)
    _sym_db.RegisterMessage(message_class)
    return message_class

  # top level enums
  for (name, enum_des) in file_des.enum_types_by_name.items():
    module['_' + name.upper()] = enum_des
    module[name] = enum_type_wrapper.EnumTypeWrapper(enum_des)
    for enum_value in enum_des.values:
      module[enum_value.name] = enum_value.number

  # top level extensions
  for (name, extension_des) in file_des.extensions_by_name.items():
    module[name.upper() + '_FIELD_NUMBER'] = extension_des.number
    module[name] = extension_des

  # services
  for (name, service) in file_des.services_by_name.items():
    module['_' + name.upper()] = service

  # Build messages.
  for (name, msg_des) in file_des.message_types_by_name.items():
    module[name] = BuildMessage(msg_des, '')


def AddHelpersToExtensions(file_des):
  """no-op to keep old generated code work with new runtime.

  Args:
    file_des: FileDescriptor of the .proto file
  """
  # TODO: Remove this on-op
  return


def BuildServices(file_des, module_name, module):
  """Builds services classes and services stub class.

  Args:
    file_des: FileDescriptor of the .proto file
    module_name: str, the name of generated _pb2 module
    module: Generated _pb2 module
  """
  # pylint: disable=g-import-not-at-top
  from google.protobuf import service_reflection
  # pylint: enable=g-import-not-at-top
  for (name, service) in file_des.services_by_name.items():
    module[name] = service_reflection.GeneratedServiceType(
        name, (),
        dict(DESCRIPTOR=service, __module__=module_name))
    stub_name = name + '_Stub'
    module[stub_name] = service_reflection.GeneratedServiceStubType(
        stub_name, (module[name],),
        dict(DESCRIPTOR=service, __module__=module_name))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_qordered_gelu.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

from logging import getLogger

from fusion_base import Fusion
from fusion_utils import FusionUtils
from onnx import helper
from onnx_model import OnnxModel

logger = getLogger(__name__)


class FusionQOrderedGelu(Fusion):
    def __init__(self, model: OnnxModel):
        super().__init__(model, "QOrderedGelu", ["Gelu", "FastGelu"])

    def fuse(self, node, input_name_to_nodes: dict, output_name_to_node: dict):
        """
        INPUT PATTERN
        Fuse (quantized) Gelu subgraph into one node QOrderedGelu:
            -> quantized input  -> DQ -> Gelu -> Q ->

        (or)

            -> quantized input  -> DQ -> FastGelu -> Q ->

        OUTPUT PATTERN
            -> QOrderedGelu ->
        """
        gelu_children = self.model.get_children(node, input_name_to_nodes)

        # Should only have 1 child - QuantizeLinear (or)
        # Should have 2 children - QuantizeLinear + Shape
        if not (
            (len(gelu_children) == 1 and gelu_children[0].op_type == "QuantizeLinear")
            or (
                len(gelu_children) == 2
                and gelu_children[0].op_type == "QuantizeLinear"
                and gelu_children[1].op_type == "Shape"
            )
        ):
            return

        downstream_quantize_node = gelu_children[0]
        downstream_shape_node = None

        if len(gelu_children) == 2:
            downstream_shape_node = gelu_children[1]

        if not FusionUtils.check_qdq_node_for_fusion(downstream_quantize_node, self.model):
            return

        # The first input to Gelu should flow through a DequantizeLinear node
        first_path_id, first_input_parent_nodes, _ = self.model.match_parent_paths(
            node,
            [(["DequantizeLinear"], [0])],
            output_name_to_node,
        )

        if first_path_id < 0:
            return

        upstream_dequantize_node = first_input_parent_nodes[0]

        if not FusionUtils.check_qdq_node_for_fusion(upstream_dequantize_node, self.model):
            return

        # Fusion logic
        subgraph_nodes = [node]  # Gelu/FastGelu
        subgraph_nodes.extend([downstream_quantize_node, upstream_dequantize_node])  # Relevant Q, DQ nodes

        if not self.model.is_safe_to_fuse_nodes(
            subgraph_nodes,
            (
                [node.output[0], downstream_quantize_node.output[0]]
                if downstream_shape_node is not None
                else downstream_quantize_node.output
            ),
            input_name_to_nodes,
            output_name_to_node,
        ):
            logger.debug("It is not safe to fuse QOrderedGelu node. Skip")
            return

        self.nodes_to_remove.extend(subgraph_nodes)

        ordered_gelu_node = helper.make_node(
            "QOrderedGelu",
            inputs=[
                upstream_dequantize_node.input[0],
                upstream_dequantize_node.input[1],
                downstream_quantize_node.input[1],
            ],
            outputs=[downstream_quantize_node.output[0]],
            name=self.model.create_node_name("QOrderedGelu", name_prefix="QOrderedGelu"),
        )

        # Arrange the downstream Shape's input to be fed from the
        # downstream QuantizeLinear node, so that fusion will
        # be deemed safe
        if downstream_shape_node is not None:
            self.model.replace_node_input(
                downstream_shape_node, downstream_shape_node.input[0], downstream_quantize_node.output[0]
            )

        # TODO: We only support CuBlasLt order ORDER_ROW for now.
        # Once we start supporting other data ordering format(s), we
        # will support user configuring the data ordering for the op.
        ordered_gelu_node.attribute.extend([helper.make_attribute("order_X", 1)])
        ordered_gelu_node.attribute.extend([helper.make_attribute("order_Y", 1)])

        ordered_gelu_node.domain = "com.microsoft"

        self.nodes_to_add.append(ordered_gelu_node)
        self.node_name_to_graph_name[ordered_gelu_node.name] = self.this_graph_name

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\window\online.py ===
from __future__ import annotations

from typing import TYPE_CHECKING

import numpy as np

from pandas.compat._optional import import_optional_dependency


def generate_online_numba_ewma_func(
    nopython: bool,
    nogil: bool,
    parallel: bool,
):
    """
    Generate a numba jitted groupby ewma function specified by values
    from engine_kwargs.

    Parameters
    ----------
    nopython : bool
        nopython to be passed into numba.jit
    nogil : bool
        nogil to be passed into numba.jit
    parallel : bool
        parallel to be passed into numba.jit

    Returns
    -------
    Numba function
    """
    if TYPE_CHECKING:
        import numba
    else:
        numba = import_optional_dependency("numba")

    @numba.jit(nopython=nopython, nogil=nogil, parallel=parallel)
    def online_ewma(
        values: np.ndarray,
        deltas: np.ndarray,
        minimum_periods: int,
        old_wt_factor: float,
        new_wt: float,
        old_wt: np.ndarray,
        adjust: bool,
        ignore_na: bool,
    ):
        """
        Compute online exponentially weighted mean per column over 2D values.

        Takes the first observation as is, then computes the subsequent
        exponentially weighted mean accounting minimum periods.
        """
        result = np.empty(values.shape)
        weighted_avg = values[0].copy()
        nobs = (~np.isnan(weighted_avg)).astype(np.int64)
        result[0] = np.where(nobs >= minimum_periods, weighted_avg, np.nan)

        for i in range(1, len(values)):
            cur = values[i]
            is_observations = ~np.isnan(cur)
            nobs += is_observations.astype(np.int64)
            for j in numba.prange(len(cur)):
                if not np.isnan(weighted_avg[j]):
                    if is_observations[j] or not ignore_na:
                        # note that len(deltas) = len(vals) - 1 and deltas[i] is to be
                        # used in conjunction with vals[i+1]
                        old_wt[j] *= old_wt_factor ** deltas[j - 1]
                        if is_observations[j]:
                            # avoid numerical errors on constant series
                            if weighted_avg[j] != cur[j]:
                                weighted_avg[j] = (
                                    (old_wt[j] * weighted_avg[j]) + (new_wt * cur[j])
                                ) / (old_wt[j] + new_wt)
                            if adjust:
                                old_wt[j] += new_wt
                            else:
                                old_wt[j] = 1.0
                elif is_observations[j]:
                    weighted_avg[j] = cur[j]

            result[i] = np.where(nobs >= minimum_periods, weighted_avg, np.nan)

        return result, old_wt

    return online_ewma


class EWMMeanState:
    def __init__(self, com, adjust, ignore_na, axis, shape) -> None:
        alpha = 1.0 / (1.0 + com)
        self.axis = axis
        self.shape = shape
        self.adjust = adjust
        self.ignore_na = ignore_na
        self.new_wt = 1.0 if adjust else alpha
        self.old_wt_factor = 1.0 - alpha
        self.old_wt = np.ones(self.shape[self.axis - 1])
        self.last_ewm = None

    def run_ewm(self, weighted_avg, deltas, min_periods, ewm_func):
        result, old_wt = ewm_func(
            weighted_avg,
            deltas,
            min_periods,
            self.old_wt_factor,
            self.new_wt,
            self.old_wt,
            self.adjust,
            self.ignore_na,
        )
        self.old_wt = old_wt
        self.last_ewm = result[-1]
        return result

    def reset(self) -> None:
        self.old_wt = np.ones(self.shape[self.axis - 1])
        self.last_ewm = None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\operations\build\wheel_legacy.py ===
import logging
import os.path
from typing import List, Optional

from pip._internal.cli.spinners import open_spinner
from pip._internal.utils.deprecation import deprecated
from pip._internal.utils.setuptools_build import make_setuptools_bdist_wheel_args
from pip._internal.utils.subprocess import call_subprocess, format_command_args

logger = logging.getLogger(__name__)


def format_command_result(
    command_args: List[str],
    command_output: str,
) -> str:
    """Format command information for logging."""
    command_desc = format_command_args(command_args)
    text = f"Command arguments: {command_desc}\n"

    if not command_output:
        text += "Command output: None"
    elif logger.getEffectiveLevel() > logging.DEBUG:
        text += "Command output: [use --verbose to show]"
    else:
        if not command_output.endswith("\n"):
            command_output += "\n"
        text += f"Command output:\n{command_output}"

    return text


def get_legacy_build_wheel_path(
    names: List[str],
    temp_dir: str,
    name: str,
    command_args: List[str],
    command_output: str,
) -> Optional[str]:
    """Return the path to the wheel in the temporary build directory."""
    # Sort for determinism.
    names = sorted(names)
    if not names:
        msg = f"Legacy build of wheel for {name!r} created no files.\n"
        msg += format_command_result(command_args, command_output)
        logger.warning(msg)
        return None

    if len(names) > 1:
        msg = (
            f"Legacy build of wheel for {name!r} created more than one file.\n"
            f"Filenames (choosing first): {names}\n"
        )
        msg += format_command_result(command_args, command_output)
        logger.warning(msg)

    return os.path.join(temp_dir, names[0])


def build_wheel_legacy(
    name: str,
    setup_py_path: str,
    source_dir: str,
    global_options: List[str],
    build_options: List[str],
    tempd: str,
) -> Optional[str]:
    """Build one unpacked package using the "legacy" build process.

    Returns path to wheel if successfully built. Otherwise, returns None.
    """
    deprecated(
        reason=(
            f"Building {name!r} using the legacy setup.py bdist_wheel mechanism, "
            "which will be removed in a future version."
        ),
        replacement=(
            "to use the standardized build interface by "
            "setting the `--use-pep517` option, "
            "(possibly combined with `--no-build-isolation`), "
            f"or adding a `pyproject.toml` file to the source tree of {name!r}"
        ),
        gone_in="25.3",
        issue=6334,
    )

    wheel_args = make_setuptools_bdist_wheel_args(
        setup_py_path,
        global_options=global_options,
        build_options=build_options,
        destination_dir=tempd,
    )

    spin_message = f"Building wheel for {name} (setup.py)"
    with open_spinner(spin_message) as spinner:
        logger.debug("Destination directory: %s", tempd)

        try:
            output = call_subprocess(
                wheel_args,
                command_desc="python setup.py bdist_wheel",
                cwd=source_dir,
                spinner=spinner,
            )
        except Exception:
            spinner.finish("error")
            logger.error("Failed building wheel for %s", name)
            return None

        names = os.listdir(tempd)
        wheel_path = get_legacy_build_wheel_path(
            names=names,
            temp_dir=tempd,
            name=name,
            command_args=wheel_args,
            command_output=output,
        )
        return wheel_path

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\expressions\funcmatrix.py ===
from .matexpr import MatrixExpr
from sympy.core.function import FunctionClass, Lambda
from sympy.core.symbol import Dummy
from sympy.core.sympify import _sympify, sympify
from sympy.matrices import Matrix
from sympy.functions.elementary.complexes import re, im


class FunctionMatrix(MatrixExpr):
    """Represents a matrix using a function (``Lambda``) which gives
    outputs according to the coordinates of each matrix entries.

    Parameters
    ==========

    rows : nonnegative integer. Can be symbolic.

    cols : nonnegative integer. Can be symbolic.

    lamda : Function, Lambda or str
        If it is a SymPy ``Function`` or ``Lambda`` instance,
        it should be able to accept two arguments which represents the
        matrix coordinates.

        If it is a pure string containing Python ``lambda`` semantics,
        it is interpreted by the SymPy parser and casted into a SymPy
        ``Lambda`` instance.

    Examples
    ========

    Creating a ``FunctionMatrix`` from ``Lambda``:

    >>> from sympy import FunctionMatrix, symbols, Lambda, MatPow
    >>> i, j, n, m = symbols('i,j,n,m')
    >>> FunctionMatrix(n, m, Lambda((i, j), i + j))
    FunctionMatrix(n, m, Lambda((i, j), i + j))

    Creating a ``FunctionMatrix`` from a SymPy function:

    >>> from sympy import KroneckerDelta
    >>> X = FunctionMatrix(3, 3, KroneckerDelta)
    >>> X.as_explicit()
    Matrix([
    [1, 0, 0],
    [0, 1, 0],
    [0, 0, 1]])

    Creating a ``FunctionMatrix`` from a SymPy undefined function:

    >>> from sympy import Function
    >>> f = Function('f')
    >>> X = FunctionMatrix(3, 3, f)
    >>> X.as_explicit()
    Matrix([
    [f(0, 0), f(0, 1), f(0, 2)],
    [f(1, 0), f(1, 1), f(1, 2)],
    [f(2, 0), f(2, 1), f(2, 2)]])

    Creating a ``FunctionMatrix`` from Python ``lambda``:

    >>> FunctionMatrix(n, m, 'lambda i, j: i + j')
    FunctionMatrix(n, m, Lambda((i, j), i + j))

    Example of lazy evaluation of matrix product:

    >>> Y = FunctionMatrix(1000, 1000, Lambda((i, j), i + j))
    >>> isinstance(Y*Y, MatPow) # this is an expression object
    True
    >>> (Y**2)[10,10] # So this is evaluated lazily
    342923500

    Notes
    =====

    This class provides an alternative way to represent an extremely
    dense matrix with entries in some form of a sequence, in a most
    sparse way.
    """
    def __new__(cls, rows, cols, lamda):
        rows, cols = _sympify(rows), _sympify(cols)
        cls._check_dim(rows)
        cls._check_dim(cols)

        lamda = sympify(lamda)
        if not isinstance(lamda, (FunctionClass, Lambda)):
            raise ValueError(
                "{} should be compatible with SymPy function classes."
                .format(lamda))

        if 2 not in lamda.nargs:
            raise ValueError(
                '{} should be able to accept 2 arguments.'.format(lamda))

        if not isinstance(lamda, Lambda):
            i, j = Dummy('i'), Dummy('j')
            lamda = Lambda((i, j), lamda(i, j))

        return super().__new__(cls, rows, cols, lamda)

    @property
    def shape(self):
        return self.args[0:2]

    @property
    def lamda(self):
        return self.args[2]

    def _entry(self, i, j, **kwargs):
        return self.lamda(i, j)

    def _eval_trace(self):
        from sympy.matrices.expressions.trace import Trace
        from sympy.concrete.summations import Sum
        return Trace(self).rewrite(Sum).doit()

    def _eval_as_real_imag(self):
        return (re(Matrix(self)), im(Matrix(self)))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\domains\field.py ===
"""Implementation of :class:`Field` class. """


from sympy.polys.domains.ring import Ring
from sympy.polys.polyerrors import NotReversible, DomainError
from sympy.utilities import public

@public
class Field(Ring):
    """Represents a field domain. """

    is_Field = True
    is_PID = True

    def get_ring(self):
        """Returns a ring associated with ``self``. """
        raise DomainError('there is no ring associated with %s' % self)

    def get_field(self):
        """Returns a field associated with ``self``. """
        return self

    def exquo(self, a, b):
        """Exact quotient of ``a`` and ``b``, implies ``__truediv__``.  """
        return a / b

    def quo(self, a, b):
        """Quotient of ``a`` and ``b``, implies ``__truediv__``. """
        return a / b

    def rem(self, a, b):
        """Remainder of ``a`` and ``b``, implies nothing.  """
        return self.zero

    def div(self, a, b):
        """Division of ``a`` and ``b``, implies ``__truediv__``. """
        return a / b, self.zero

    def gcd(self, a, b):
        """
        Returns GCD of ``a`` and ``b``.

        This definition of GCD over fields allows to clear denominators
        in `primitive()`.

        Examples
        ========

        >>> from sympy.polys.domains import QQ
        >>> from sympy import S, gcd, primitive
        >>> from sympy.abc import x

        >>> QQ.gcd(QQ(2, 3), QQ(4, 9))
        2/9
        >>> gcd(S(2)/3, S(4)/9)
        2/9
        >>> primitive(2*x/3 + S(4)/9)
        (2/9, 3*x + 2)

        """
        try:
            ring = self.get_ring()
        except DomainError:
            return self.one

        p = ring.gcd(self.numer(a), self.numer(b))
        q = ring.lcm(self.denom(a), self.denom(b))

        return self.convert(p, ring)/q

    def gcdex(self, a, b):
        """
        Returns x, y, g such that a * x + b * y == g == gcd(a, b)
        """
        d = self.gcd(a, b)

        if a == self.zero:
            if b == self.zero:
                return self.zero, self.one, self.zero
            else:
                return self.zero, d/b, d
        else:
            return d/a, self.zero, d

    def lcm(self, a, b):
        """
        Returns LCM of ``a`` and ``b``.

        >>> from sympy.polys.domains import QQ
        >>> from sympy import S, lcm

        >>> QQ.lcm(QQ(2, 3), QQ(4, 9))
        4/3
        >>> lcm(S(2)/3, S(4)/9)
        4/3

        """

        try:
            ring = self.get_ring()
        except DomainError:
            return a*b

        p = ring.lcm(self.numer(a), self.numer(b))
        q = ring.gcd(self.denom(a), self.denom(b))

        return self.convert(p, ring)/q

    def revert(self, a):
        """Returns ``a**(-1)`` if possible. """
        if a:
            return 1/a
        else:
            raise NotReversible('zero is not reversible')

    def is_unit(self, a):
        """Return true if ``a`` is a invertible"""
        return bool(a)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\domains\ring.py ===
"""Implementation of :class:`Ring` class. """


from sympy.polys.domains.domain import Domain
from sympy.polys.polyerrors import ExactQuotientFailed, NotInvertible, NotReversible

from sympy.utilities import public

@public
class Ring(Domain):
    """Represents a ring domain. """

    is_Ring = True

    def get_ring(self):
        """Returns a ring associated with ``self``. """
        return self

    def exquo(self, a, b):
        """Exact quotient of ``a`` and ``b``, implies ``__floordiv__``.  """
        if a % b:
            raise ExactQuotientFailed(a, b, self)
        else:
            return a // b

    def quo(self, a, b):
        """Quotient of ``a`` and ``b``, implies ``__floordiv__``. """
        return a // b

    def rem(self, a, b):
        """Remainder of ``a`` and ``b``, implies ``__mod__``.  """
        return a % b

    def div(self, a, b):
        """Division of ``a`` and ``b``, implies ``__divmod__``. """
        return divmod(a, b)

    def invert(self, a, b):
        """Returns inversion of ``a mod b``. """
        s, t, h = self.gcdex(a, b)

        if self.is_one(h):
            return s % b
        else:
            raise NotInvertible("zero divisor")

    def revert(self, a):
        """Returns ``a**(-1)`` if possible. """
        if self.is_one(a) or self.is_one(-a):
            return a
        else:
            raise NotReversible('only units are reversible in a ring')

    def is_unit(self, a):
        try:
            self.revert(a)
            return True
        except NotReversible:
            return False

    def numer(self, a):
        """Returns numerator of ``a``. """
        return a

    def denom(self, a):
        """Returns denominator of `a`. """
        return self.one

    def free_module(self, rank):
        """
        Generate a free module of rank ``rank`` over self.

        >>> from sympy.abc import x
        >>> from sympy import QQ
        >>> QQ.old_poly_ring(x).free_module(2)
        QQ[x]**2
        """
        raise NotImplementedError

    def ideal(self, *gens):
        """
        Generate an ideal of ``self``.

        >>> from sympy.abc import x
        >>> from sympy import QQ
        >>> QQ.old_poly_ring(x).ideal(x**2)
        <x**2>
        """
        from sympy.polys.agca.ideals import ModuleImplementedIdeal
        return ModuleImplementedIdeal(self, self.free_module(1).submodule(
            *[[x] for x in gens]))

    def quotient_ring(self, e):
        """
        Form a quotient ring of ``self``.

        Here ``e`` can be an ideal or an iterable.

        >>> from sympy.abc import x
        >>> from sympy import QQ
        >>> QQ.old_poly_ring(x).quotient_ring(QQ.old_poly_ring(x).ideal(x**2))
        QQ[x]/<x**2>
        >>> QQ.old_poly_ring(x).quotient_ring([x**2])
        QQ[x]/<x**2>

        The division operator has been overloaded for this:

        >>> QQ.old_poly_ring(x)/[x**2]
        QQ[x]/<x**2>
        """
        from sympy.polys.agca.ideals import Ideal
        from sympy.polys.domains.quotientring import QuotientRing
        if not isinstance(e, Ideal):
            e = self.ideal(*e)
        return QuotientRing(self, e)

    def __truediv__(self, e):
        return self.quotient_ring(e)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_core\_tests\test_local.py ===
import pytest

from trio import run
from trio.lowlevel import RunVar, RunVarToken

from ... import _core


# scary runvar tests
def test_runvar_smoketest() -> None:
    t1 = RunVar[str]("test1")
    t2 = RunVar[str]("test2", default="catfish")

    assert repr(t1) == "<RunVar name='test1'>"

    async def first_check() -> None:
        with pytest.raises(LookupError):
            t1.get()

        t1.set("swordfish")
        assert t1.get() == "swordfish"
        assert t2.get() == "catfish"
        assert t2.get(default="eel") == "eel"

        t2.set("goldfish")
        assert t2.get() == "goldfish"
        assert t2.get(default="tuna") == "goldfish"

    async def second_check() -> None:
        with pytest.raises(LookupError):
            t1.get()

        assert t2.get() == "catfish"

    run(first_check)
    run(second_check)


def test_runvar_resetting() -> None:
    t1 = RunVar[str]("test1")
    t2 = RunVar[str]("test2", default="dogfish")
    t3 = RunVar[str]("test3")

    async def reset_check() -> None:
        token = t1.set("moonfish")
        assert t1.get() == "moonfish"
        t1.reset(token)

        with pytest.raises(TypeError):
            t1.reset(None)  # type: ignore[arg-type]

        with pytest.raises(LookupError):
            t1.get()

        token2 = t2.set("catdogfish")
        assert t2.get() == "catdogfish"
        t2.reset(token2)
        assert t2.get() == "dogfish"

        with pytest.raises(ValueError, match=r"^token has already been used$"):
            t2.reset(token2)

        token3 = t3.set("basculin")
        assert t3.get() == "basculin"

        with pytest.raises(ValueError, match=r"^token is not for us$"):
            t1.reset(token3)

    run(reset_check)


def test_runvar_sync() -> None:
    t1 = RunVar[str]("test1")

    async def sync_check() -> None:
        async def task1() -> None:
            t1.set("plaice")
            assert t1.get() == "plaice"

        async def task2(tok: RunVarToken[str]) -> None:
            t1.reset(tok)

            with pytest.raises(LookupError):
                t1.get()

            t1.set("haddock")

        async with _core.open_nursery() as n:
            token = t1.set("cod")
            assert t1.get() == "cod"

            n.start_soon(task1)
            await _core.wait_all_tasks_blocked()
            assert t1.get() == "plaice"

            n.start_soon(task2, token)
            await _core.wait_all_tasks_blocked()
            assert t1.get() == "haddock"

    run(sync_check)


def test_accessing_runvar_outside_run_call_fails() -> None:
    t1 = RunVar[str]("test1")

    with pytest.raises(RuntimeError):
        t1.set("asdf")

    with pytest.raises(RuntimeError):
        t1.get()

    async def get_token() -> RunVarToken[str]:
        return t1.set("ok")

    token = run(get_token)

    with pytest.raises(RuntimeError):
        t1.reset(token)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\tests\test_matrixbase.py ===
import concurrent.futures
import random
from collections.abc import Hashable

from sympy import (
    Abs, Add, Array, DeferredVector, E, Expr, FiniteSet, Float, Function,
    GramSchmidt, I, ImmutableDenseMatrix, ImmutableMatrix,
    ImmutableSparseMatrix, Integer, KroneckerDelta, MatPow, Matrix,
    MatrixSymbol, Max, Min, MutableDenseMatrix, MutableSparseMatrix, Poly, Pow,
    PurePoly, Q, Quaternion, Rational, RootOf, S, SparseMatrix, Symbol, Tuple,
    Wild, banded, casoratian, cos, diag, diff, exp, expand, eye, floor, hessian,
    integrate, log, matrix_multiply_elementwise, nan, ones, oo, pi, randMatrix,
    rot_axis1, rot_axis2, rot_axis3, rot_ccw_axis1, rot_ccw_axis2,
    rot_ccw_axis3, signsimp, simplify, sin, sqrt, sstr, symbols, sympify, tan,
    trigsimp, wronskian, zeros, cancel)
from sympy.abc import a, b, c, d, t, x, y, z
from sympy.core.kind import NumberKind, UndefinedKind
from sympy.matrices.determinant import _find_reasonable_pivot_naive
from sympy.matrices.exceptions import (
    MatrixError, NonSquareMatrixError, ShapeError)
from sympy.matrices.kind import MatrixKind
from sympy.matrices.utilities import _dotprodsimp_state, _simplify, dotprodsimp
from sympy.tensor.array.array_derivatives import ArrayDerivative
from sympy.testing.pytest import (
    ignore_warnings, raises, skip, skip_under_pyodide, slow,
    warns_deprecated_sympy)
from sympy.utilities.iterables import capture, iterable
from importlib.metadata import version

all_classes = (Matrix, SparseMatrix, ImmutableMatrix, ImmutableSparseMatrix)
mutable_classes = (Matrix, SparseMatrix)
immutable_classes = (ImmutableMatrix, ImmutableSparseMatrix)


def test__MinimalMatrix():
    x = Matrix(2, 3, [1, 2, 3, 4, 5, 6])
    assert x.rows == 2
    assert x.cols == 3
    assert x[2] == 3
    assert x[1, 1] == 5
    assert list(x) == [1, 2, 3, 4, 5, 6]
    assert list(x[1, :]) == [4, 5, 6]
    assert list(x[:, 1]) == [2, 5]
    assert list(x[:, :]) == list(x)
    assert x[:, :] == x
    assert Matrix(x) == x
    assert Matrix([[1, 2, 3], [4, 5, 6]]) == x
    assert Matrix(([1, 2, 3], [4, 5, 6])) == x
    assert Matrix([(1, 2, 3), (4, 5, 6)]) == x
    assert Matrix(((1, 2, 3), (4, 5, 6))) == x
    assert not (Matrix([[1, 2], [3, 4], [5, 6]]) == x)


def test_kind():
    assert Matrix([[1, 2], [3, 4]]).kind == MatrixKind(NumberKind)
    assert Matrix([[0, 0], [0, 0]]).kind == MatrixKind(NumberKind)
    assert Matrix(0, 0, []).kind == MatrixKind(NumberKind)
    assert Matrix([[x]]).kind == MatrixKind(NumberKind)
    assert Matrix([[1, Matrix([[1]])]]).kind == MatrixKind(UndefinedKind)
    assert SparseMatrix([[1]]).kind == MatrixKind(NumberKind)
    assert SparseMatrix([[1, Matrix([[1]])]]).kind == MatrixKind(UndefinedKind)


def test_todok():
    a, b, c, d = symbols('a:d')
    m1 = MutableDenseMatrix([[a, b], [c, d]])
    m2 = ImmutableDenseMatrix([[a, b], [c, d]])
    m3 = MutableSparseMatrix([[a, b], [c, d]])
    m4 = ImmutableSparseMatrix([[a, b], [c, d]])
    assert m1.todok() == m2.todok() == m3.todok() == m4.todok() == \
        {(0, 0): a, (0, 1): b, (1, 0): c, (1, 1): d}


def test_tolist():
    lst = [[S.One, S.Half, x*y, S.Zero], [x, y, z, x**2], [y, -S.One, z*x, 3]]
    flat_lst = [S.One, S.Half, x*y, S.Zero, x, y, z, x**2, y, -S.One, z*x, 3]
    m = Matrix(3, 4, flat_lst)
    assert m.tolist() == lst


def test_todod():
    m = Matrix([[S.One, 0], [0, S.Half], [x, 0]])
    dict = {0: {0: S.One}, 1: {1: S.Half}, 2: {0: x}}
    assert m.todod() == dict


def test_row_col_del():
    e = ImmutableMatrix(3, 3, [1, 2, 3, 4, 5, 6, 7, 8, 9])
    raises(IndexError, lambda: e.row_del(5))
    raises(IndexError, lambda: e.row_del(-5))
    raises(IndexError, lambda: e.col_del(5))
    raises(IndexError, lambda: e.col_del(-5))

    assert e.row_del(2) == e.row_del(-1) == Matrix([[1, 2, 3], [4, 5, 6]])
    assert e.col_del(2) == e.col_del(-1) == Matrix([[1, 2], [4, 5], [7, 8]])

    assert e.row_del(1) == e.row_del(-2) == Matrix([[1, 2, 3], [7, 8, 9]])
    assert e.col_del(1) == e.col_del(-2) == Matrix([[1, 3], [4, 6], [7, 9]])


def test_get_diag_blocks1():
    a = Matrix([[1, 2], [2, 3]])
    b = Matrix([[3, x], [y, 3]])
    c = Matrix([[3, x, 3], [y, 3, z], [x, y, z]])
    assert a.get_diag_blocks() == [a]
    assert b.get_diag_blocks() == [b]
    assert c.get_diag_blocks() == [c]


def test_get_diag_blocks2():
    a = Matrix([[1, 2], [2, 3]])
    b = Matrix([[3, x], [y, 3]])
    c = Matrix([[3, x, 3], [y, 3, z], [x, y, z]])
    A, B, C, D = diag(a, b, b), diag(a, b, c), diag(a, c, b), diag(c, c, b)
    A = Matrix(A.rows, A.cols, A)
    B = Matrix(B.rows, B.cols, B)
    C = Matrix(C.rows, C.cols, C)
    D = Matrix(D.rows, D.cols, D)

    assert A.get_diag_blocks() == [a, b, b]
    assert B.get_diag_blocks() == [a, b, c]
    assert C.get_diag_blocks() == [a, c, b]
    assert D.get_diag_blocks() == [c, c, b]


def test_row_col():
    m = Matrix(3, 3, [1, 2, 3, 4, 5, 6, 7, 8, 9])
    assert m.row(0) == Matrix(1, 3, [1, 2, 3])
    assert m.col(0) == Matrix(3, 1, [1, 4, 7])


def test_row_join():
    assert eye(3).row_join(Matrix([7, 7, 7])) == \
           Matrix([[1, 0, 0, 7],
                   [0, 1, 0, 7],
                   [0, 0, 1, 7]])


def test_col_join():
    assert eye(3).col_join(Matrix([[7, 7, 7]])) == \
           Matrix([[1, 0, 0],
                   [0, 1, 0],
                   [0, 0, 1],
                   [7, 7, 7]])


def test_row_insert():
    r4 = Matrix([[4, 4, 4]])
    for i in range(-4, 5):
        l = [1, 0, 0]
        l.insert(i, 4)
        assert eye(3).row_insert(i, r4).col(0).flat() == l


def test_col_insert():
    c4 = Matrix([4, 4, 4])
    for i in range(-4, 5):
        l = [0, 0, 0]
        l.insert(i, 4)
        assert zeros(3).col_insert(i, c4).row(0).flat() == l
    # issue 13643
    assert eye(6).col_insert(3, Matrix([[2, 2], [2, 2], [2, 2], [2, 2], [2, 2], [2, 2]])) == \
           Matrix([[1, 0, 0, 2, 2, 0, 0, 0],
                   [0, 1, 0, 2, 2, 0, 0, 0],
                   [0, 0, 1, 2, 2, 0, 0, 0],
                   [0, 0, 0, 2, 2, 1, 0, 0],
                   [0, 0, 0, 2, 2, 0, 1, 0],
                   [0, 0, 0, 2, 2, 0, 0, 1]])


def test_extract():
    m = Matrix(4, 3, lambda i, j: i*3 + j)
    assert m.extract([0, 1, 3], [0, 1]) == Matrix(3, 2, [0, 1, 3, 4, 9, 10])
    assert m.extract([0, 3], [0, 0, 2]) == Matrix(2, 3, [0, 0, 2, 9, 9, 11])
    assert m.extract(range(4), range(3)) == m
    raises(IndexError, lambda: m.extract([4], [0]))
    raises(IndexError, lambda: m.extract([0], [3]))


def test_hstack():
    m = Matrix(4, 3, lambda i, j: i*3 + j)
    m2 = Matrix(3, 4, lambda i, j: i*3 + j)
    assert m == m.hstack(m)
    assert m.hstack(m, m, m) == Matrix.hstack(m, m, m) == Matrix([
                [0,  1,  2, 0,  1,  2, 0,  1,  2],
                [3,  4,  5, 3,  4,  5, 3,  4,  5],
                [6,  7,  8, 6,  7,  8, 6,  7,  8],
                [9, 10, 11, 9, 10, 11, 9, 10, 11]])
    raises(ShapeError, lambda: m.hstack(m, m2))
    assert Matrix.hstack() == Matrix()

    # test regression #12938
    M1 = Matrix.zeros(0, 0)
    M2 = Matrix.zeros(0, 1)
    M3 = Matrix.zeros(0, 2)
    M4 = Matrix.zeros(0, 3)
    m = Matrix.hstack(M1, M2, M3, M4)
    assert m.rows == 0 and m.cols == 6


def test_vstack():
    m = Matrix(4, 3, lambda i, j: i*3 + j)
    m2 = Matrix(3, 4, lambda i, j: i*3 + j)
    assert m == m.vstack(m)
    assert m.vstack(m, m, m) == Matrix.vstack(m, m, m) == Matrix([
                                [0,  1,  2],
                                [3,  4,  5],
                                [6,  7,  8],
                                [9, 10, 11],
                                [0,  1,  2],
                                [3,  4,  5],
                                [6,  7,  8],
                                [9, 10, 11],
                                [0,  1,  2],
                                [3,  4,  5],
                                [6,  7,  8],
                                [9, 10, 11]])
    raises(ShapeError, lambda: m.vstack(m, m2))
    assert Matrix.vstack() == Matrix()


def test_has():
    A = Matrix(((x, y), (2, 3)))
    assert A.has(x)
    assert not A.has(z)
    assert A.has(Symbol)

    A = Matrix(((2, y), (2, 3)))
    assert not A.has(x)


def test_is_anti_symmetric():
    x = symbols('x')
    assert Matrix(2, 1, [1, 2]).is_anti_symmetric() is False
    m = Matrix(3, 3, [0, x**2 + 2*x + 1, y, -(x + 1)**2, 0, x*y, -y, -x*y, 0])
    assert m.is_anti_symmetric() is True
    assert m.is_anti_symmetric(simplify=False) is None
    assert m.is_anti_symmetric(simplify=lambda x: x) is None

    m = Matrix(3, 3, [x.expand() for x in m])
    assert m.is_anti_symmetric(simplify=False) is True
    m = Matrix(3, 3, [x.expand() for x in [S.One] + list(m)[1:]])
    assert m.is_anti_symmetric() is False


def test_is_hermitian():
    a = Matrix([[1, I], [-I, 1]])
    assert a.is_hermitian
    a = Matrix([[2*I, I], [-I, 1]])
    assert a.is_hermitian is False
    a = Matrix([[x, I], [-I, 1]])
    assert a.is_hermitian is None
    a = Matrix([[x, 1], [-I, 1]])
    assert a.is_hermitian is False


def test_is_symbolic():
    a = Matrix([[x, x], [x, x]])
    assert a.is_symbolic() is True
    a = Matrix([[1, 2, 3, 4], [5, 6, 7, 8]])
    assert a.is_symbolic() is False
    a = Matrix([[1, 2, 3, 4], [5, 6, x, 8]])
    assert a.is_symbolic() is True
    a = Matrix([[1, x, 3]])
    assert a.is_symbolic() is True
    a = Matrix([[1, 2, 3]])
    assert a.is_symbolic() is False
    a = Matrix([[1], [x], [3]])
    assert a.is_symbolic() is True
    a = Matrix([[1], [2], [3]])
    assert a.is_symbolic() is False


def test_is_square():
    m = Matrix([[1], [1]])
    m2 = Matrix([[2, 2], [2, 2]])
    assert not m.is_square
    assert m2.is_square


def test_is_symmetric():
    m = Matrix(2, 2, [0, 1, 1, 0])
    assert m.is_symmetric()
    m = Matrix(2, 2, [0, 1, 0, 1])
    assert not m.is_symmetric()


def test_is_hessenberg():
    A = Matrix([[3, 4, 1], [2, 4, 5], [0, 1, 2]])
    assert A.is_upper_hessenberg
    A = Matrix(3, 3, [3, 2, 0, 4, 4, 1, 1, 5, 2])
    assert A.is_lower_hessenberg
    A = Matrix(3, 3, [3, 2, -1, 4, 4, 1, 1, 5, 2])
    assert A.is_lower_hessenberg is False
    assert A.is_upper_hessenberg is False

    A = Matrix([[3, 4, 1], [2, 4, 5], [3, 1, 2]])
    assert not A.is_upper_hessenberg


def test_values():
    assert set(Matrix(2, 2, [0, 1, 2, 3]
        ).values()) == {1, 2, 3}
    x = Symbol('x', real=True)
    assert set(Matrix(2, 2, [x, 0, 0, 1]
        ).values()) == {x, 1}


def test_conjugate():
    M = Matrix([[0, I, 5],
                [1, 2, 0]])

    assert M.T == Matrix([[0, 1],
                          [I, 2],
                          [5, 0]])

    assert M.C == Matrix([[0, -I, 5],
                          [1,  2, 0]])
    assert M.C == M.conjugate()

    assert M.H == M.T.C
    assert M.H == Matrix([[ 0, 1],
                          [-I, 2],
                          [ 5, 0]])


def test_doit():
    a = Matrix([[Add(x, x, evaluate=False)]])
    assert a[0] != 2*x
    assert a.doit() == Matrix([[2*x]])


def test_evalf():
    a = Matrix(2, 1, [sqrt(5), 6])
    assert all(a.evalf()[i] == a[i].evalf() for i in range(2))
    assert all(a.evalf(2)[i] == a[i].evalf(2) for i in range(2))
    assert all(a.n(2)[i] == a[i].n(2) for i in range(2))


def test_replace():
    F, G = symbols('F, G', cls=Function)
    K = Matrix(2, 2, lambda i, j: G(i+j))
    M = Matrix(2, 2, lambda i, j: F(i+j))
    N = M.replace(F, G)
    assert N == K


def test_replace_map():
    F, G = symbols('F, G', cls=Function)
    M = Matrix(2, 2, lambda i, j: F(i+j))
    N, d = M.replace(F, G, True)
    assert N == Matrix(2, 2, lambda i, j: G(i+j))
    assert d == {F(0): G(0), F(1): G(1), F(2): G(2)}

def test_numpy_conversion():
    try:
        from numpy import array, array_equal
    except ImportError:
        skip('NumPy must be available to test creating matrices from ndarrays')
    A = Matrix([[1,2], [3,4]])
    np_array = array([[1,2], [3,4]])
    assert array_equal(array(A), np_array)
    assert array_equal(array(A, copy=True), np_array)
    if(int(version('numpy').split('.')[0]) >= 2): #run this test only if numpy is new enough that copy variable is passed properly.
        raises(TypeError, lambda: array(A, copy=False))

def test_rot90():
    A = Matrix([[1, 2], [3, 4]])
    assert A == A.rot90(0) == A.rot90(4)
    assert A.rot90(2) == A.rot90(-2) == A.rot90(6) == Matrix(((4, 3), (2, 1)))
    assert A.rot90(3) == A.rot90(-1) == A.rot90(7) == Matrix(((2, 4), (1, 3)))
    assert A.rot90() == A.rot90(-7) == A.rot90(-3) == Matrix(((3, 1), (4, 2)))


def test_subs():
    assert Matrix([[1, x], [x, 4]]).subs(x, 5) == Matrix([[1, 5], [5, 4]])
    assert Matrix([[x, 2], [x + y, 4]]).subs([[x, -1], [y, -2]]) == \
           Matrix([[-1, 2], [-3, 4]])
    assert Matrix([[x, 2], [x + y, 4]]).subs([(x, -1), (y, -2)]) == \
           Matrix([[-1, 2], [-3, 4]])
    assert Matrix([[x, 2], [x + y, 4]]).subs({x: -1, y: -2}) == \
           Matrix([[-1, 2], [-3, 4]])
    assert Matrix([[x*y]]).subs({x: y - 1, y: x - 1}, simultaneous=True) == \
           Matrix([[(x - 1)*(y - 1)]])


def test_permute():
    a = Matrix(3, 4, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12])

    raises(IndexError, lambda: a.permute([[0, 5]]))
    raises(ValueError, lambda: a.permute(Symbol('x')))
    b = a.permute_rows([[0, 2], [0, 1]])
    assert a.permute([[0, 2], [0, 1]]) == b == Matrix([
                                            [5,  6,  7,  8],
                                            [9, 10, 11, 12],
                                            [1,  2,  3,  4]])

    b = a.permute_cols([[0, 2], [0, 1]])
    assert a.permute([[0, 2], [0, 1]], orientation='cols') == b ==\
                            Matrix([
                            [ 2,  3, 1,  4],
                            [ 6,  7, 5,  8],
                            [10, 11, 9, 12]])

    b = a.permute_cols([[0, 2], [0, 1]], direction='backward')
    assert a.permute([[0, 2], [0, 1]], orientation='cols', direction='backward') == b ==\
                            Matrix([
                            [ 3, 1,  2,  4],
                            [ 7, 5,  6,  8],
                            [11, 9, 10, 12]])

    assert a.permute([1, 2, 0, 3]) == Matrix([
                                            [5,  6,  7,  8],
                                            [9, 10, 11, 12],
                                            [1,  2,  3,  4]])

    from sympy.combinatorics import Permutation
    assert a.permute(Permutation([1, 2, 0, 3])) == Matrix([
                                            [5,  6,  7,  8],
                                            [9, 10, 11, 12],
                                            [1,  2,  3,  4]])

def test_upper_triangular():

    A = Matrix([
                [1, 1, 1, 1],
                [1, 1, 1, 1],
                [1, 1, 1, 1],
                [1, 1, 1, 1]
            ])

    R = A.upper_triangular(2)
    assert R == Matrix([
                        [0, 0, 1, 1],
                        [0, 0, 0, 1],
                        [0, 0, 0, 0],
                        [0, 0, 0, 0]
                    ])

    R = A.upper_triangular(-2)
    assert R == Matrix([
                        [1, 1, 1, 1],
                        [1, 1, 1, 1],
                        [1, 1, 1, 1],
                        [0, 1, 1, 1]
                    ])

    R = A.upper_triangular()
    assert R == Matrix([
                        [1, 1, 1, 1],
                        [0, 1, 1, 1],
                        [0, 0, 1, 1],
                        [0, 0, 0, 1]
                    ])


def test_lower_triangular():
    A = Matrix([
                        [1, 1, 1, 1],
                        [1, 1, 1, 1],
                        [1, 1, 1, 1],
                        [1, 1, 1, 1]
                    ])

    L = A.lower_triangular()
    assert L == Matrix([
                        [1, 0, 0, 0],
                        [1, 1, 0, 0],
                        [1, 1, 1, 0],
                        [1, 1, 1, 1]])

    L = A.lower_triangular(2)
    assert L == Matrix([
                        [1, 1, 1, 0],
                        [1, 1, 1, 1],
                        [1, 1, 1, 1],
                        [1, 1, 1, 1]
                    ])

    L = A.lower_triangular(-2)
    assert L == Matrix([
                        [0, 0, 0, 0],
                        [0, 0, 0, 0],
                        [1, 0, 0, 0],
                        [1, 1, 0, 0]
                    ])


def test_add():
    m = Matrix([[1, 2, 3], [x, y, x], [2*y, -50, z*x]])
    assert m + m == Matrix([[2, 4, 6], [2*x, 2*y, 2*x], [4*y, -100, 2*z*x]])
    n = Matrix(1, 2, [1, 2])
    raises(ShapeError, lambda: m + n)


def test_matmul():
    a = Matrix([[1, 2], [3, 4]])

    assert a.__matmul__(2) == NotImplemented

    assert a.__rmatmul__(2) == NotImplemented

    #This is done this way because @ is only supported in Python 3.5+
    #To check 2@a case
    try:
        eval('2 @ a')
    except SyntaxError:
        pass
    except TypeError:  #TypeError is raised in case of NotImplemented is returned
        pass

    #Check a@2 case
    try:
        eval('a @ 2')
    except SyntaxError:
        pass
    except TypeError:  #TypeError is raised in case of NotImplemented is returned
        pass


def test_non_matmul():
    """
    Test that if explicitly specified as non-matrix, mul reverts
    to scalar multiplication.
    """
    class foo(Expr):
        is_Matrix=False
        is_MatrixLike=False
        shape = (1, 1)

    A = Matrix([[1, 2], [3, 4]])
    b = foo()
    assert b*A == Matrix([[b, 2*b], [3*b, 4*b]])
    assert A*b == Matrix([[b, 2*b], [3*b, 4*b]])


def test_neg():
    n = Matrix(1, 2, [1, 2])
    assert -n == Matrix(1, 2, [-1, -2])


def test_sub():
    n = Matrix(1, 2, [1, 2])
    assert n - n == Matrix(1, 2, [0, 0])


def test_div():
    n = Matrix(1, 2, [1, 2])
    assert n/2 == Matrix(1, 2, [S.Half, S(2)/2])


def test_eye():
    assert list(Matrix.eye(2, 2)) == [1, 0, 0, 1]
    assert list(Matrix.eye(2)) == [1, 0, 0, 1]
    assert type(Matrix.eye(2)) == Matrix
    assert type(Matrix.eye(2, cls=Matrix)) == Matrix


def test_ones():
    assert list(Matrix.ones(2, 2)) == [1, 1, 1, 1]
    assert list(Matrix.ones(2)) == [1, 1, 1, 1]
    assert Matrix.ones(2, 3) == Matrix([[1, 1, 1], [1, 1, 1]])
    assert type(Matrix.ones(2)) == Matrix
    assert type(Matrix.ones(2, cls=Matrix)) == Matrix


def test_zeros():
    assert list(Matrix.zeros(2, 2)) == [0, 0, 0, 0]
    assert list(Matrix.zeros(2)) == [0, 0, 0, 0]
    assert Matrix.zeros(2, 3) == Matrix([[0, 0, 0], [0, 0, 0]])
    assert type(Matrix.zeros(2)) == Matrix
    assert type(Matrix.zeros(2, cls=Matrix)) == Matrix


def test_diag_make():
    diag = Matrix.diag
    a = Matrix([[1, 2], [2, 3]])
    b = Matrix([[3, x], [y, 3]])
    c = Matrix([[3, x, 3], [y, 3, z], [x, y, z]])
    assert diag(a, b, b) == Matrix([
        [1, 2, 0, 0, 0, 0],
        [2, 3, 0, 0, 0, 0],
        [0, 0, 3, x, 0, 0],
        [0, 0, y, 3, 0, 0],
        [0, 0, 0, 0, 3, x],
        [0, 0, 0, 0, y, 3],
    ])
    assert diag(a, b, c) == Matrix([
        [1, 2, 0, 0, 0, 0, 0],
        [2, 3, 0, 0, 0, 0, 0],
        [0, 0, 3, x, 0, 0, 0],
        [0, 0, y, 3, 0, 0, 0],
        [0, 0, 0, 0, 3, x, 3],
        [0, 0, 0, 0, y, 3, z],
        [0, 0, 0, 0, x, y, z],
    ])
    assert diag(a, c, b) == Matrix([
        [1, 2, 0, 0, 0, 0, 0],
        [2, 3, 0, 0, 0, 0, 0],
        [0, 0, 3, x, 3, 0, 0],
        [0, 0, y, 3, z, 0, 0],
        [0, 0, x, y, z, 0, 0],
        [0, 0, 0, 0, 0, 3, x],
        [0, 0, 0, 0, 0, y, 3],
    ])
    a = Matrix([x, y, z])
    b = Matrix([[1, 2], [3, 4]])
    c = Matrix([[5, 6]])
    # this "wandering diagonal" is what makes this
    # a block diagonal where each block is independent
    # of the others
    assert diag(a, 7, b, c) == Matrix([
        [x, 0, 0, 0, 0, 0],
        [y, 0, 0, 0, 0, 0],
        [z, 0, 0, 0, 0, 0],
        [0, 7, 0, 0, 0, 0],
        [0, 0, 1, 2, 0, 0],
        [0, 0, 3, 4, 0, 0],
        [0, 0, 0, 0, 5, 6]])
    raises(ValueError, lambda: diag(a, 7, b, c, rows=5))
    assert diag(1) == Matrix([[1]])
    assert diag(1, rows=2) == Matrix([[1, 0], [0, 0]])
    assert diag(1, cols=2) == Matrix([[1, 0], [0, 0]])
    assert diag(1, rows=3, cols=2) == Matrix([[1, 0], [0, 0], [0, 0]])
    assert diag(*[2, 3]) == Matrix([
        [2, 0],
        [0, 3]])
    assert diag(Matrix([2, 3])) == Matrix([
        [2],
        [3]])
    assert diag([1, [2, 3], 4], unpack=False) == \
            diag([[1], [2, 3], [4]], unpack=False) == Matrix([
        [1, 0],
        [2, 3],
        [4, 0]])
    assert type(diag(1)) == Matrix
    assert type(diag(1, cls=Matrix)) == Matrix
    assert Matrix.diag([1, 2, 3]) == Matrix.diag(1, 2, 3)
    assert Matrix.diag([1, 2, 3], unpack=False).shape == (3, 1)
    assert Matrix.diag([[1, 2, 3]]).shape == (3, 1)
    assert Matrix.diag([[1, 2, 3]], unpack=False).shape == (1, 3)
    assert Matrix.diag([[[1, 2, 3]]]).shape == (1, 3)
    # kerning can be used to move the starting point
    assert Matrix.diag(ones(0, 2), 1, 2) == Matrix([
        [0, 0, 1, 0],
        [0, 0, 0, 2]])
    assert Matrix.diag(ones(2, 0), 1, 2) == Matrix([
        [0, 0],
        [0, 0],
        [1, 0],
        [0, 2]])


def test_diagonal():
    m = Matrix(3, 3, range(9))
    d = m.diagonal()
    assert d == m.diagonal(0)
    assert tuple(d) == (0, 4, 8)
    assert tuple(m.diagonal(1)) == (1, 5)
    assert tuple(m.diagonal(-1)) == (3, 7)
    assert tuple(m.diagonal(2)) == (2,)
    assert type(m.diagonal()) == type(m)
    s = SparseMatrix(3, 3, {(1, 1): 1})
    assert type(s.diagonal()) == type(s)
    assert type(m) != type(s)
    raises(ValueError, lambda: m.diagonal(3))
    raises(ValueError, lambda: m.diagonal(-3))
    raises(ValueError, lambda: m.diagonal(pi))
    M = ones(2, 3)
    assert banded({i: list(M.diagonal(i))
        for i in range(1-M.rows, M.cols)}) == M


def test_jordan_block():
    assert Matrix.jordan_block(3, 2) == Matrix.jordan_block(3, eigenvalue=2) \
            == Matrix.jordan_block(size=3, eigenvalue=2) \
            == Matrix.jordan_block(3, 2, band='upper') \
            == Matrix.jordan_block(
                size=3, eigenval=2, eigenvalue=2) \
            == Matrix([
                [2, 1, 0],
                [0, 2, 1],
                [0, 0, 2]])

    assert Matrix.jordan_block(3, 2, band='lower') == Matrix([
                    [2, 0, 0],
                    [1, 2, 0],
                    [0, 1, 2]])
    # missing eigenvalue
    raises(ValueError, lambda: Matrix.jordan_block(2))
    # non-integral size
    raises(ValueError, lambda: Matrix.jordan_block(3.5, 2))
    # size not specified
    raises(ValueError, lambda: Matrix.jordan_block(eigenvalue=2))
    # inconsistent eigenvalue
    raises(ValueError,
    lambda: Matrix.jordan_block(
        eigenvalue=2, eigenval=4))

    # Using alias keyword
    assert Matrix.jordan_block(size=3, eigenvalue=2) == \
        Matrix.jordan_block(size=3, eigenval=2)


def test_orthogonalize():
    m = Matrix([[1, 2], [3, 4]])
    assert m.orthogonalize(Matrix([[2], [1]])) == [Matrix([[2], [1]])]
    assert m.orthogonalize(Matrix([[2], [1]]), normalize=True) == \
        [Matrix([[2*sqrt(5)/5], [sqrt(5)/5]])]
    assert m.orthogonalize(Matrix([[1], [2]]), Matrix([[-1], [4]])) == \
        [Matrix([[1], [2]]), Matrix([[Rational(-12, 5)], [Rational(6, 5)]])]
    assert m.orthogonalize(Matrix([[0], [0]]), Matrix([[-1], [4]])) == \
        [Matrix([[-1], [4]])]
    assert m.orthogonalize(Matrix([[0], [0]])) == []

    n = Matrix([[9, 1, 9], [3, 6, 10], [8, 5, 2]])
    vecs = [Matrix([[-5], [1]]), Matrix([[-5], [2]]), Matrix([[-5], [-2]])]
    assert n.orthogonalize(*vecs) == \
        [Matrix([[-5], [1]]), Matrix([[Rational(5, 26)], [Rational(25, 26)]])]

    vecs = [Matrix([0, 0, 0]), Matrix([1, 2, 3]), Matrix([1, 4, 5])]
    raises(ValueError, lambda: Matrix.orthogonalize(*vecs, rankcheck=True))

    vecs = [Matrix([1, 2, 3]), Matrix([4, 5, 6]), Matrix([7, 8, 9])]
    raises(ValueError, lambda: Matrix.orthogonalize(*vecs, rankcheck=True))

def test_wilkinson():

    wminus, wplus = Matrix.wilkinson(1)
    assert wminus == Matrix([
                                [-1, 1, 0],
                                [1, 0, 1],
                                [0, 1, 1]])
    assert wplus == Matrix([
                            [1, 1, 0],
                            [1, 0, 1],
                            [0, 1, 1]])

    wminus, wplus = Matrix.wilkinson(3)
    assert wminus == Matrix([
                                [-3,  1,  0, 0, 0, 0, 0],
                                [1, -2,  1, 0, 0, 0, 0],
                                [0,  1, -1, 1, 0, 0, 0],
                                [0,  0,  1, 0, 1, 0, 0],
                                [0,  0,  0, 1, 1, 1, 0],
                                [0,  0,  0, 0, 1, 2, 1],

      [0,  0,  0, 0, 0, 1, 3]])

    assert wplus == Matrix([
                            [3, 1, 0, 0, 0, 0, 0],
                            [1, 2, 1, 0, 0, 0, 0],
                            [0, 1, 1, 1, 0, 0, 0],
                            [0, 0, 1, 0, 1, 0, 0],
                            [0, 0, 0, 1, 1, 1, 0],
                            [0, 0, 0, 0, 1, 2, 1],
                            [0, 0, 0, 0, 0, 1, 3]])


def test_limit():
    x, y = symbols('x y')
    m = Matrix(2, 1, [1/x, y])
    assert m.limit(x, 5) == Matrix(2, 1, [Rational(1, 5), y])
    A = Matrix(((1, 4, sin(x)/x), (y, 2, 4), (10, 5, x**2 + 1)))
    assert A.limit(x, 0) == Matrix(((1, 4, 1), (y, 2, 4), (10, 5, 1)))


def test_issue_13774():
    M = Matrix([[1, 2, 3], [4, 5, 6], [7, 8, 9]])
    v = [1, 1, 1]
    raises(TypeError, lambda: M*v)
    raises(TypeError, lambda: v*M)


def test_companion():
    x = Symbol('x')
    y = Symbol('y')
    raises(ValueError, lambda: Matrix.companion(1))
    raises(ValueError, lambda: Matrix.companion(Poly([1], x)))
    raises(ValueError, lambda: Matrix.companion(Poly([2, 1], x)))
    raises(ValueError, lambda: Matrix.companion(Poly(x*y, [x, y])))

    c0, c1, c2 = symbols('c0:3')
    assert Matrix.companion(Poly([1, c0], x)) == Matrix([-c0])
    assert Matrix.companion(Poly([1, c1, c0], x)) == \
        Matrix([[0, -c0], [1, -c1]])
    assert Matrix.companion(Poly([1, c2, c1, c0], x)) == \
        Matrix([[0, 0, -c0], [1, 0, -c1], [0, 1, -c2]])


def test_issue_10589():
    x, y, z = symbols("x, y z")
    M1 = Matrix([x, y, z])
    M1 = M1.subs(zip([x, y, z], [1, 2, 3]))
    assert M1 == Matrix([[1], [2], [3]])

    M2 = Matrix([[x, x, x, x, x], [x, x, x, x, x], [x, x, x, x, x]])
    M2 = M2.subs(zip([x], [1]))
    assert M2 == Matrix([[1, 1, 1, 1, 1], [1, 1, 1, 1, 1], [1, 1, 1, 1, 1]])


def test_rmul_pr19860():
    class Foo(ImmutableDenseMatrix):
        _op_priority = MutableDenseMatrix._op_priority + 0.01

    a = Matrix(2, 2, [1, 2, 3, 4])
    b = Foo(2, 2, [1, 2, 3, 4])

    # This would throw a RecursionError: maximum recursion depth
    # since b always has higher priority even after a.as_mutable()
    c = a*b

    assert isinstance(c, Foo)
    assert c == Matrix([[7, 10], [15, 22]])


def test_issue_18956():
    A = Array([[1, 2], [3, 4]])
    B = Matrix([[1,2],[3,4]])
    raises(TypeError, lambda: B + A)
    raises(TypeError, lambda: A + B)


def test__eq__():
    class My(object):
        def __iter__(self):
            yield 1
            yield 2
            return
        def __getitem__(self, i):
            return list(self)[i]
    a = Matrix(2, 1, [1, 2])
    assert a != My()
    class My_sympy(My):
        def _sympy_(self):
            return Matrix(self)
    assert a == My_sympy()


def test_args():
    for n, cls in enumerate(all_classes):
        m = cls.zeros(3, 2)
        # all should give back the same type of arguments, e.g. ints for shape
        assert m.shape == (3, 2) and all(type(i) is int for i in m.shape)
        assert m.rows == 3 and type(m.rows) is int
        assert m.cols == 2 and type(m.cols) is int
        if not n % 2:
            assert type(m.flat()) in (list, tuple, Tuple)
        else:
            assert type(m.todok()) is dict


def test_deprecated_mat_smat():
    for cls in Matrix, ImmutableMatrix:
        m = cls.zeros(3, 2)
        with warns_deprecated_sympy():
            mat = m._mat
        assert mat == m.flat()
    for cls in SparseMatrix, ImmutableSparseMatrix:
        m = cls.zeros(3, 2)
        with warns_deprecated_sympy():
            smat = m._smat
        assert smat == m.todok()


def test_division():
    v = Matrix(1, 2, [x, y])
    assert v/z == Matrix(1, 2, [x/z, y/z])


def test_sum():
    m = Matrix([[1, 2, 3], [x, y, x], [2*y, -50, z*x]])
    assert m + m == Matrix([[2, 4, 6], [2*x, 2*y, 2*x], [4*y, -100, 2*z*x]])
    n = Matrix(1, 2, [1, 2])
    raises(ShapeError, lambda: m + n)


def test_abs():
    m = Matrix([[1, -2], [x, y]])
    assert abs(m) == Matrix([[1, 2], [Abs(x), Abs(y)]])
    m = Matrix(1, 2, [-3, x])
    n = Matrix(1, 2, [3, Abs(x)])
    assert abs(m) == n


def test_addition():
    a = Matrix((
        (1, 2),
        (3, 1),
    ))

    b = Matrix((
        (1, 2),
        (3, 0),
    ))

    assert a + b == a.add(b) == Matrix([[2, 4], [6, 1]])


def test_fancy_index_matrix():
    for M in (Matrix, SparseMatrix):
        a = M(3, 3, range(9))
        assert a == a[:, :]
        assert a[1, :] == Matrix(1, 3, [3, 4, 5])
        assert a[:, 1] == Matrix([1, 4, 7])
        assert a[[0, 1], :] == Matrix([[0, 1, 2], [3, 4, 5]])
        assert a[[0, 1], 2] == a[[0, 1], [2]]
        assert a[2, [0, 1]] == a[[2], [0, 1]]
        assert a[:, [0, 1]] == Matrix([[0, 1], [3, 4], [6, 7]])
        assert a[0, 0] == 0
        assert a[0:2, :] == Matrix([[0, 1, 2], [3, 4, 5]])
        assert a[:, 0:2] == Matrix([[0, 1], [3, 4], [6, 7]])
        assert a[::2, 1] == a[[0, 2], 1]
        assert a[1, ::2] == a[1, [0, 2]]
        a = M(3, 3, range(9))
        assert a[[0, 2, 1, 2, 1], :] == Matrix([
            [0, 1, 2],
            [6, 7, 8],
            [3, 4, 5],
            [6, 7, 8],
            [3, 4, 5]])
        assert a[:, [0,2,1,2,1]] == Matrix([
            [0, 2, 1, 2, 1],
            [3, 5, 4, 5, 4],
            [6, 8, 7, 8, 7]])

    a = SparseMatrix.zeros(3)
    a[1, 2] = 2
    a[0, 1] = 3
    a[2, 0] = 4
    assert a.extract([1, 1], [2]) == Matrix([
    [2],
    [2]])
    assert a.extract([1, 0], [2, 2, 2]) == Matrix([
    [2, 2, 2],
    [0, 0, 0]])
    assert a.extract([1, 0, 1, 2], [2, 0, 1, 0]) == Matrix([
        [2, 0, 0, 0],
        [0, 0, 3, 0],
        [2, 0, 0, 0],
        [0, 4, 0, 4]])


def test_multiplication():
    a = Matrix((
        (1, 2),
        (3, 1),
        (0, 6),
    ))

    b = Matrix((
        (1, 2),
        (3, 0),
    ))

    raises(ShapeError, lambda: b*a)
    raises(TypeError, lambda: a*{})

    c = a*b
    assert c[0, 0] == 7
    assert c[0, 1] == 2
    assert c[1, 0] == 6
    assert c[1, 1] == 6
    assert c[2, 0] == 18
    assert c[2, 1] == 0

    c = a @ b
    assert c[0, 0] == 7
    assert c[0, 1] == 2
    assert c[1, 0] == 6
    assert c[1, 1] == 6
    assert c[2, 0] == 18
    assert c[2, 1] == 0

    h = matrix_multiply_elementwise(a, c)
    assert h == a.multiply_elementwise(c)
    assert h[0, 0] == 7
    assert h[0, 1] == 4
    assert h[1, 0] == 18
    assert h[1, 1] == 6
    assert h[2, 0] == 0
    assert h[2, 1] == 0
    raises(ShapeError, lambda: matrix_multiply_elementwise(a, b))

    c = b * Symbol("x")
    assert isinstance(c, Matrix)
    assert c[0, 0] == x
    assert c[0, 1] == 2*x
    assert c[1, 0] == 3*x
    assert c[1, 1] == 0

    c2 = x * b
    assert c == c2

    c = 5 * b
    assert isinstance(c, Matrix)
    assert c[0, 0] == 5
    assert c[0, 1] == 2*5
    assert c[1, 0] == 3*5
    assert c[1, 1] == 0

    M = Matrix([[oo, 0], [0, oo]])
    assert M ** 2 == M

    M = Matrix([[oo, oo], [0, 0]])
    assert M ** 2 == Matrix([[nan, nan], [nan, nan]])

    # https://github.com/sympy/sympy/issues/22353
    A = Matrix(ones(3, 1))
    _h = -Rational(1, 2)
    B = Matrix([_h, _h, _h])
    assert A.multiply_elementwise(B) == Matrix([
        [_h],
        [_h],
        [_h]])


def test_power():
    raises(NonSquareMatrixError, lambda: Matrix((1, 2))**2)

    A = Matrix([[2, 3], [4, 5]])
    assert A**5 == Matrix([[6140, 8097], [10796, 14237]])
    A = Matrix([[2, 1, 3], [4, 2, 4], [6, 12, 1]])
    assert A**3 == Matrix([[290, 262, 251], [448, 440, 368], [702, 954, 433]])
    assert A**0 == eye(3)
    assert A**1 == A
    assert (Matrix([[2]]) ** 100)[0, 0] == 2**100
    assert Matrix([[1, 2], [3, 4]])**Integer(2) == Matrix([[7, 10], [15, 22]])
    A = Matrix([[1,2],[4,5]])
    assert A.pow(20, method='cayley') == A.pow(20, method='multiply')
    assert A**Integer(2) == Matrix([[9, 12], [24, 33]])
    assert eye(2)**10000000 == eye(2)

    A = Matrix([[33, 24], [48, 57]])
    assert (A**S.Half)[:] == [5, 2, 4, 7]
    A = Matrix([[0, 4], [-1, 5]])
    assert (A**S.Half)**2 == A

    assert Matrix([[1, 0], [1, 1]])**S.Half == Matrix([[1, 0], [S.Half, 1]])
    assert Matrix([[1, 0], [1, 1]])**0.5 == Matrix([[1, 0], [0.5, 1]])
    from sympy.abc import n
    assert Matrix([[1, a], [0, 1]])**n == Matrix([[1, a*n], [0, 1]])
    assert Matrix([[b, a], [0, b]])**n == Matrix([[b**n, a*b**(n-1)*n], [0, b**n]])
    assert Matrix([
        [a**n, a**(n - 1)*n, (a**n*n**2 - a**n*n)/(2*a**2)],
        [   0,         a**n,                  a**(n - 1)*n],
        [   0,            0,                          a**n]])
    assert Matrix([[a, 1, 0], [0, a, 0], [0, 0, b]])**n == Matrix([
        [a**n, a**(n-1)*n, 0],
        [0, a**n, 0],
        [0, 0, b**n]])

    A = Matrix([[1, 0], [1, 7]])
    assert A._matrix_pow_by_jordan_blocks(S(3)) == A._eval_pow_by_recursion(3)
    A = Matrix([[2]])
    assert A**10 == Matrix([[2**10]]) == A._matrix_pow_by_jordan_blocks(S(10)) == \
        A._eval_pow_by_recursion(10)

    # testing a matrix that cannot be jordan blocked issue 11766
    m = Matrix([[3, 0, 0, 0, -3], [0, -3, -3, 0, 3], [0, 3, 0, 3, 0], [0, 0, 3, 0, 3], [3, 0, 0, 3, 0]])
    raises(MatrixError, lambda: m._matrix_pow_by_jordan_blocks(S(10)))

    # test issue 11964
    raises(MatrixError, lambda: Matrix([[1, 1], [3, 3]])._matrix_pow_by_jordan_blocks(S(-10)))
    A = Matrix([[0, 1, 0], [0, 0, 1], [0, 0, 0]])  # Nilpotent jordan block size 3
    assert A**10.0 == Matrix([[0, 0, 0], [0, 0, 0], [0, 0, 0]])
    raises(ValueError, lambda: A**2.1)
    raises(ValueError, lambda: A**Rational(3, 2))
    A = Matrix([[8, 1], [3, 2]])
    assert A**10.0 == Matrix([[1760744107, 272388050], [817164150, 126415807]])
    A = Matrix([[0, 0, 1], [0, 0, 1], [0, 0, 1]])  # Nilpotent jordan block size 1
    assert A**10.0 == Matrix([[0, 0, 1], [0, 0, 1], [0, 0, 1]])
    A = Matrix([[0, 1, 0], [0, 0, 1], [0, 0, 1]])  # Nilpotent jordan block size 2
    assert A**10.0 == Matrix([[0, 0, 1], [0, 0, 1], [0, 0, 1]])
    n = Symbol('n', integer=True)
    assert isinstance(A**n, MatPow)
    n = Symbol('n', integer=True, negative=True)
    raises(ValueError, lambda: A**n)
    n = Symbol('n', integer=True, nonnegative=True)
    assert A**n == Matrix([
        [KroneckerDelta(0, n), KroneckerDelta(1, n), -KroneckerDelta(0, n) - KroneckerDelta(1, n) + 1],
        [                   0, KroneckerDelta(0, n),                         1 - KroneckerDelta(0, n)],
        [                   0,                    0,                                                1]])
    assert A**(n + 2) == Matrix([[0, 0, 1], [0, 0, 1], [0, 0, 1]])
    raises(ValueError, lambda: A**Rational(3, 2))
    A = Matrix([[0, 0, 1], [3, 0, 1], [4, 3, 1]])
    assert A**5.0 == Matrix([[168,  72,  89], [291, 144, 161], [572, 267, 329]])
    assert A**5.0 == A**5
    A = Matrix([[0, 1, 0],[-1, 0, 0],[0, 0, 0]])
    n = Symbol("n")
    An = A**n
    assert An.subs(n, 2).doit() == A**2
    raises(ValueError, lambda: An.subs(n, -2).doit())
    assert An * An == A**(2*n)

    # concretizing behavior for non-integer and complex powers
    A = Matrix([[0,0,0],[0,0,0],[0,0,0]])
    n = Symbol('n', integer=True, positive=True)
    assert A**n == A
    n = Symbol('n', integer=True, nonnegative=True)
    assert A**n == diag(0**n, 0**n, 0**n)
    assert (A**n).subs(n, 0) == eye(3)
    assert (A**n).subs(n, 1) == zeros(3)
    A = Matrix ([[2,0,0],[0,2,0],[0,0,2]])
    assert A**2.1 == diag (2**2.1, 2**2.1, 2**2.1)
    assert A**I == diag (2**I, 2**I, 2**I)
    A = Matrix([[0, 1, 0], [0, 0, 1], [0, 0, 1]])
    raises(ValueError, lambda: A**2.1)
    raises(ValueError, lambda: A**I)
    A = Matrix([[S.Half, S.Half], [S.Half, S.Half]])
    assert A**S.Half == A
    A = Matrix([[1, 1],[3, 3]])
    assert A**S.Half == Matrix ([[S.Half, S.Half], [3*S.Half, 3*S.Half]])


def test_issue_17247_expression_blowup_1():
    M = Matrix([[1+x, 1-x], [1-x, 1+x]])
    with dotprodsimp(True):
        assert M.exp().expand() == Matrix([
            [ (exp(2*x) + exp(2))/2, (-exp(2*x) + exp(2))/2],
            [(-exp(2*x) + exp(2))/2,  (exp(2*x) + exp(2))/2]])


def test_issue_17247_expression_blowup_2():
    M = Matrix([[1+x, 1-x], [1-x, 1+x]])
    with dotprodsimp(True):
        P, J = M.jordan_form ()
        assert P*J*P.inv()


def test_issue_17247_expression_blowup_3():
    M = Matrix([[1+x, 1-x], [1-x, 1+x]])
    with dotprodsimp(True):
        assert M**100 == Matrix([
            [633825300114114700748351602688*x**100 + 633825300114114700748351602688, 633825300114114700748351602688 - 633825300114114700748351602688*x**100],
            [633825300114114700748351602688 - 633825300114114700748351602688*x**100, 633825300114114700748351602688*x**100 + 633825300114114700748351602688]])


def test_issue_17247_expression_blowup_4():
# This matrix takes extremely long on current master even with intermediate simplification so an abbreviated version is used. It is left here for test in case of future optimizations.
#     M = Matrix(S('''[
#         [             -3/4,       45/32 - 37*I/16,         1/4 + I/2,      -129/64 - 9*I/64,      1/4 - 5*I/16,      65/128 + 87*I/64,         -9/32 - I/16,      183/256 - 97*I/128,       3/64 + 13*I/64,         -23/32 - 59*I/256,      15/128 - 3*I/32,        19/256 + 551*I/1024],
#         [-149/64 + 49*I/32, -177/128 - 1369*I/128,  125/64 + 87*I/64, -2063/256 + 541*I/128,  85/256 - 33*I/16,  805/128 + 2415*I/512, -219/128 + 115*I/256, 6301/4096 - 6609*I/1024,  119/128 + 143*I/128, -10879/2048 + 4343*I/4096,  129/256 - 549*I/512, 42533/16384 + 29103*I/8192],
#         [          1/2 - I,         9/4 + 55*I/16,              -3/4,       45/32 - 37*I/16,         1/4 + I/2,      -129/64 - 9*I/64,         1/4 - 5*I/16,        65/128 + 87*I/64,         -9/32 - I/16,        183/256 - 97*I/128,       3/64 + 13*I/64,          -23/32 - 59*I/256],
#         [   -5/8 - 39*I/16,   2473/256 + 137*I/64, -149/64 + 49*I/32, -177/128 - 1369*I/128,  125/64 + 87*I/64, -2063/256 + 541*I/128,     85/256 - 33*I/16,    805/128 + 2415*I/512, -219/128 + 115*I/256,   6301/4096 - 6609*I/1024,  119/128 + 143*I/128,  -10879/2048 + 4343*I/4096],
#         [            1 + I,         -19/4 + 5*I/4,           1/2 - I,         9/4 + 55*I/16,              -3/4,       45/32 - 37*I/16,            1/4 + I/2,        -129/64 - 9*I/64,         1/4 - 5*I/16,          65/128 + 87*I/64,         -9/32 - I/16,         183/256 - 97*I/128],
#         [         21/8 + I,    -537/64 + 143*I/16,    -5/8 - 39*I/16,   2473/256 + 137*I/64, -149/64 + 49*I/32, -177/128 - 1369*I/128,     125/64 + 87*I/64,   -2063/256 + 541*I/128,     85/256 - 33*I/16,      805/128 + 2415*I/512, -219/128 + 115*I/256,    6301/4096 - 6609*I/1024],
#         [               -2,         17/4 - 13*I/2,             1 + I,         -19/4 + 5*I/4,           1/2 - I,         9/4 + 55*I/16,                 -3/4,         45/32 - 37*I/16,            1/4 + I/2,          -129/64 - 9*I/64,         1/4 - 5*I/16,           65/128 + 87*I/64],
#         [     1/4 + 13*I/4,    -825/64 - 147*I/32,          21/8 + I,    -537/64 + 143*I/16,    -5/8 - 39*I/16,   2473/256 + 137*I/64,    -149/64 + 49*I/32,   -177/128 - 1369*I/128,     125/64 + 87*I/64,     -2063/256 + 541*I/128,     85/256 - 33*I/16,       805/128 + 2415*I/512],
#         [             -4*I,            27/2 + 6*I,                -2,         17/4 - 13*I/2,             1 + I,         -19/4 + 5*I/4,              1/2 - I,           9/4 + 55*I/16,                 -3/4,           45/32 - 37*I/16,            1/4 + I/2,           -129/64 - 9*I/64],
#         [      1/4 + 5*I/2,       -23/8 - 57*I/16,      1/4 + 13*I/4,    -825/64 - 147*I/32,          21/8 + I,    -537/64 + 143*I/16,       -5/8 - 39*I/16,     2473/256 + 137*I/64,    -149/64 + 49*I/32,     -177/128 - 1369*I/128,     125/64 + 87*I/64,      -2063/256 + 541*I/128],
#         [               -4,               9 - 5*I,              -4*I,            27/2 + 6*I,                -2,         17/4 - 13*I/2,                1 + I,           -19/4 + 5*I/4,              1/2 - I,             9/4 + 55*I/16,                 -3/4,            45/32 - 37*I/16],
#         [             -2*I,        119/8 + 29*I/4,       1/4 + 5*I/2,       -23/8 - 57*I/16,      1/4 + 13*I/4,    -825/64 - 147*I/32,             21/8 + I,      -537/64 + 143*I/16,       -5/8 - 39*I/16,       2473/256 + 137*I/64,    -149/64 + 49*I/32,      -177/128 - 1369*I/128]]'''))
#     assert M**10 == Matrix([
#         [    7*(-221393644768594642173548179825793834595 - 1861633166167425978847110897013541127952*I)/9671406556917033397649408,      15*(31670992489131684885307005100073928751695 + 10329090958303458811115024718207404523808*I)/77371252455336267181195264,   7*(-3710978679372178839237291049477017392703 + 1377706064483132637295566581525806894169*I)/19342813113834066795298816,            (9727707023582419994616144751727760051598 - 59261571067013123836477348473611225724433*I)/9671406556917033397649408,      (31896723509506857062605551443641668183707 + 54643444538699269118869436271152084599580*I)/38685626227668133590597632,       (-2024044860947539028275487595741003997397402 + 130959428791783397562960461903698670485863*I)/309485009821345068724781056,     3*(26190251453797590396533756519358368860907 - 27221191754180839338002754608545400941638*I)/77371252455336267181195264,      (1154643595139959842768960128434994698330461 + 3385496216250226964322872072260446072295634*I)/618970019642690137449562112,     3*(-31849347263064464698310044805285774295286 - 11877437776464148281991240541742691164309*I)/77371252455336267181195264,     (4661330392283532534549306589669150228040221 - 4171259766019818631067810706563064103956871*I)/1237940039285380274899124224,        (9598353794289061833850770474812760144506 + 358027153990999990968244906482319780943983*I)/309485009821345068724781056,     (-9755135335127734571547571921702373498554177 - 4837981372692695195747379349593041939686540*I)/2475880078570760549798248448],
#         [(-379516731607474268954110071392894274962069 - 422272153179747548473724096872271700878296*I)/77371252455336267181195264, (41324748029613152354787280677832014263339501 - 12715121258662668420833935373453570749288074*I)/1237940039285380274899124224, (-339216903907423793947110742819264306542397 + 494174755147303922029979279454787373566517*I)/77371252455336267181195264, (-18121350839962855576667529908850640619878381 - 37413012454129786092962531597292531089199003*I)/1237940039285380274899124224, (2489661087330511608618880408199633556675926 + 1137821536550153872137379935240732287260863*I)/309485009821345068724781056, (-136644109701594123227587016790354220062972119 + 110130123468183660555391413889600443583585272*I)/4951760157141521099596496896, (1488043981274920070468141664150073426459593 - 9691968079933445130866371609614474474327650*I)/1237940039285380274899124224,  27*(4636797403026872518131756991410164760195942 + 3369103221138229204457272860484005850416533*I)/4951760157141521099596496896, (-8534279107365915284081669381642269800472363 + 2241118846262661434336333368511372725482742*I)/1237940039285380274899124224,  (60923350128174260992536531692058086830950875 - 263673488093551053385865699805250505661590126*I)/9903520314283042199192993792, (18520943561240714459282253753348921824172569 + 24846649186468656345966986622110971925703604*I)/4951760157141521099596496896,  (-232781130692604829085973604213529649638644431 + 35981505277760667933017117949103953338570617*I)/9903520314283042199192993792],
#         [      (8742968295129404279528270438201520488950 + 3061473358639249112126847237482570858327*I)/4835703278458516698824704,      (-245657313712011778432792959787098074935273 + 253113767861878869678042729088355086740856*I)/38685626227668133590597632,      (1947031161734702327107371192008011621193 - 19462330079296259148177542369999791122762*I)/9671406556917033397649408,        (552856485625209001527688949522750288619217 + 392928441196156725372494335248099016686580*I)/77371252455336267181195264,      (-44542866621905323121630214897126343414629 + 3265340021421335059323962377647649632959*I)/19342813113834066795298816,          (136272594005759723105646069956434264218730 - 330975364731707309489523680957584684763587*I)/38685626227668133590597632,       (27392593965554149283318732469825168894401 + 75157071243800133880129376047131061115278*I)/38685626227668133590597632,      7*(-357821652913266734749960136017214096276154 - 45509144466378076475315751988405961498243*I)/309485009821345068724781056,       (104485001373574280824835174390219397141149 - 99041000529599568255829489765415726168162*I)/77371252455336267181195264,      (1198066993119982409323525798509037696321291 + 4249784165667887866939369628840569844519936*I)/618970019642690137449562112,       (-114985392587849953209115599084503853611014 - 52510376847189529234864487459476242883449*I)/77371252455336267181195264,      (6094620517051332877965959223269600650951573 - 4683469779240530439185019982269137976201163*I)/1237940039285380274899124224],
#         [ (611292255597977285752123848828590587708323 - 216821743518546668382662964473055912169502*I)/77371252455336267181195264,  (-1144023204575811464652692396337616594307487 + 12295317806312398617498029126807758490062855*I)/309485009821345068724781056, (-374093027769390002505693378578475235158281 - 573533923565898290299607461660384634333639*I)/77371252455336267181195264,   (47405570632186659000138546955372796986832987 - 2837476058950808941605000274055970055096534*I)/1237940039285380274899124224,   (-571573207393621076306216726219753090535121 + 533381457185823100878764749236639320783831*I)/77371252455336267181195264,     (-7096548151856165056213543560958582513797519 - 24035731898756040059329175131592138642195366*I)/618970019642690137449562112,  (2396762128833271142000266170154694033849225 + 1448501087375679588770230529017516492953051*I)/309485009821345068724781056, (-150609293845161968447166237242456473262037053 + 92581148080922977153207018003184520294188436*I)/4951760157141521099596496896, 5*(270278244730804315149356082977618054486347 - 1997830155222496880429743815321662710091562*I)/1237940039285380274899124224,   (62978424789588828258068912690172109324360330 + 44803641177219298311493356929537007630129097*I)/2475880078570760549798248448, 19*(-451431106327656743945775812536216598712236 + 114924966793632084379437683991151177407937*I)/1237940039285380274899124224,   (63417747628891221594106738815256002143915995 - 261508229397507037136324178612212080871150958*I)/9903520314283042199192993792],
#         [     (-2144231934021288786200752920446633703357 + 2305614436009705803670842248131563850246*I)/1208925819614629174706176,       (-90720949337459896266067589013987007078153 - 221951119475096403601562347412753844534569*I)/19342813113834066795298816,      (11590973613116630788176337262688659880376 + 6514520676308992726483494976339330626159*I)/4835703278458516698824704,      3*(-131776217149000326618649542018343107657237 + 79095042939612668486212006406818285287004*I)/38685626227668133590597632,       (10100577916793945997239221374025741184951 - 28631383488085522003281589065994018550748*I)/9671406556917033397649408,         67*(10090295594251078955008130473573667572549 + 10449901522697161049513326446427839676762*I)/77371252455336267181195264,       (-54270981296988368730689531355811033930513 - 3413683117592637309471893510944045467443*I)/19342813113834066795298816,         (440372322928679910536575560069973699181278 - 736603803202303189048085196176918214409081*I)/77371252455336267181195264,        (33220374714789391132887731139763250155295 + 92055083048787219934030779066298919603554*I)/38685626227668133590597632,      5*(-594638554579967244348856981610805281527116 - 82309245323128933521987392165716076704057*I)/309485009821345068724781056,       (128056368815300084550013708313312073721955 - 114619107488668120303579745393765245911404*I)/77371252455336267181195264,       21*(59839959255173222962789517794121843393573 + 241507883613676387255359616163487405826334*I)/618970019642690137449562112],
#         [ (-13454485022325376674626653802541391955147 + 184471402121905621396582628515905949793486*I)/19342813113834066795298816,   (-6158730123400322562149780662133074862437105 - 3416173052604643794120262081623703514107476*I)/154742504910672534362390528,  (770558003844914708453618983120686116100419 - 127758381209767638635199674005029818518766*I)/77371252455336267181195264,   (-4693005771813492267479835161596671660631703 + 12703585094750991389845384539501921531449948*I)/309485009821345068724781056,   (-295028157441149027913545676461260860036601 - 841544569970643160358138082317324743450770*I)/77371252455336267181195264,     (56716442796929448856312202561538574275502893 + 7216818824772560379753073185990186711454778*I)/1237940039285380274899124224,  15*(-87061038932753366532685677510172566368387 + 61306141156647596310941396434445461895538*I)/154742504910672534362390528,    (-3455315109680781412178133042301025723909347 - 24969329563196972466388460746447646686670670*I)/618970019642690137449562112,   (2453418854160886481106557323699250865361849 + 1497886802326243014471854112161398141242514*I)/309485009821345068724781056, (-151343224544252091980004429001205664193082173 + 90471883264187337053549090899816228846836628*I)/4951760157141521099596496896,   (1652018205533026103358164026239417416432989 - 9959733619236515024261775397109724431400162*I)/1237940039285380274899124224,  3*(40676374242956907656984876692623172736522006 + 31023357083037817469535762230872667581366205*I)/4951760157141521099596496896],
#         [     (-1226990509403328460274658603410696548387 - 4131739423109992672186585941938392788458*I)/1208925819614629174706176,         (162392818524418973411975140074368079662703 + 23706194236915374831230612374344230400704*I)/9671406556917033397649408,      (-3935678233089814180000602553655565621193 + 2283744757287145199688061892165659502483*I)/1208925819614629174706176,         (-2400210250844254483454290806930306285131 - 315571356806370996069052930302295432758205*I)/19342813113834066795298816,       (13365917938215281056563183751673390817910 + 15911483133819801118348625831132324863881*I)/4835703278458516698824704,        3*(-215950551370668982657516660700301003897855 + 51684341999223632631602864028309400489378*I)/38685626227668133590597632,        (20886089946811765149439844691320027184765 - 30806277083146786592790625980769214361844*I)/9671406556917033397649408,        (562180634592713285745940856221105667874855 + 1031543963988260765153550559766662245114916*I)/77371252455336267181195264,       (-65820625814810177122941758625652476012867 - 12429918324787060890804395323920477537595*I)/19342813113834066795298816,         (319147848192012911298771180196635859221089 - 402403304933906769233365689834404519960394*I)/38685626227668133590597632,        (23035615120921026080284733394359587955057 + 115351677687031786114651452775242461310624*I)/38685626227668133590597632,      (-3426830634881892756966440108592579264936130 - 1022954961164128745603407283836365128598559*I)/309485009821345068724781056],
#         [ (-192574788060137531023716449082856117537757 - 69222967328876859586831013062387845780692*I)/19342813113834066795298816,     (2736383768828013152914815341491629299773262 - 2773252698016291897599353862072533475408743*I)/77371252455336267181195264,  (-23280005281223837717773057436155921656805 + 214784953368021840006305033048142888879224*I)/19342813113834066795298816,     (-3035247484028969580570400133318947903462326 - 2195168903335435855621328554626336958674325*I)/77371252455336267181195264,     (984552428291526892214541708637840971548653 - 64006622534521425620714598573494988589378*I)/77371252455336267181195264,      (-3070650452470333005276715136041262898509903 + 7286424705750810474140953092161794621989080*I)/154742504910672534362390528,    (-147848877109756404594659513386972921139270 - 416306113044186424749331418059456047650861*I)/38685626227668133590597632,    (55272118474097814260289392337160619494260781 + 7494019668394781211907115583302403519488058*I)/1237940039285380274899124224,     (-581537886583682322424771088996959213068864 + 542191617758465339135308203815256798407429*I)/77371252455336267181195264,    (-6422548983676355789975736799494791970390991 - 23524183982209004826464749309156698827737702*I)/618970019642690137449562112,     7*(180747195387024536886923192475064903482083 + 84352527693562434817771649853047924991804*I)/154742504910672534362390528, (-135485179036717001055310712747643466592387031 + 102346575226653028836678855697782273460527608*I)/4951760157141521099596496896],
#         [        (3384238362616083147067025892852431152105 + 156724444932584900214919898954874618256*I)/604462909807314587353088,        (-59558300950677430189587207338385764871866 + 114427143574375271097298201388331237478857*I)/4835703278458516698824704,      (-1356835789870635633517710130971800616227 - 7023484098542340388800213478357340875410*I)/1208925819614629174706176,          (234884918567993750975181728413524549575881 + 79757294640629983786895695752733890213506*I)/9671406556917033397649408,        (-7632732774935120473359202657160313866419 + 2905452608512927560554702228553291839465*I)/1208925819614629174706176,           (52291747908702842344842889809762246649489 - 520996778817151392090736149644507525892649*I)/19342813113834066795298816,        (17472406829219127839967951180375981717322 + 23464704213841582137898905375041819568669*I)/4835703278458516698824704,        (-911026971811893092350229536132730760943307 + 150799318130900944080399439626714846752360*I)/38685626227668133590597632,         (26234457233977042811089020440646443590687 - 45650293039576452023692126463683727692890*I)/9671406556917033397649408,       3*(288348388717468992528382586652654351121357 + 454526517721403048270274049572136109264668*I)/77371252455336267181195264,        (-91583492367747094223295011999405657956347 - 12704691128268298435362255538069612411331*I)/19342813113834066795298816,          (411208730251327843849027957710164064354221 - 569898526380691606955496789378230959965898*I)/38685626227668133590597632],
#         [    (27127513117071487872628354831658811211795 - 37765296987901990355760582016892124833857*I)/4835703278458516698824704,     (1741779916057680444272938534338833170625435 + 3083041729779495966997526404685535449810378*I)/77371252455336267181195264, 3*(-60642236251815783728374561836962709533401 - 24630301165439580049891518846174101510744*I)/19342813113834066795298816,      3*(445885207364591681637745678755008757483408 - 350948497734812895032502179455610024541643*I)/38685626227668133590597632,    (-47373295621391195484367368282471381775684 + 219122969294089357477027867028071400054973*I)/19342813113834066795298816,       (-2801565819673198722993348253876353741520438 - 2250142129822658548391697042460298703335701*I)/77371252455336267181195264,      (801448252275607253266997552356128790317119 - 50890367688077858227059515894356594900558*I)/77371252455336267181195264,    (-5082187758525931944557763799137987573501207 + 11610432359082071866576699236013484487676124*I)/309485009821345068724781056,     (-328925127096560623794883760398247685166830 - 643447969697471610060622160899409680422019*I)/77371252455336267181195264,    15*(2954944669454003684028194956846659916299765 + 33434406416888505837444969347824812608566*I)/1237940039285380274899124224,      (-415749104352001509942256567958449835766827 + 479330966144175743357171151440020955412219*I)/77371252455336267181195264,  3*(-4639987285852134369449873547637372282914255 - 11994411888966030153196659207284951579243273*I)/1237940039285380274899124224],
#         [       (-478846096206269117345024348666145495601 + 1249092488629201351470551186322814883283*I)/302231454903657293676544,         (-17749319421930878799354766626365926894989 - 18264580106418628161818752318217357231971*I)/1208925819614629174706176,         (2801110795431528876849623279389579072819 + 363258850073786330770713557775566973248*I)/604462909807314587353088,          (-59053496693129013745775512127095650616252 + 78143588734197260279248498898321500167517*I)/4835703278458516698824704,         (-283186724922498212468162690097101115349 - 6443437753863179883794497936345437398276*I)/1208925819614629174706176,            (188799118826748909206887165661384998787543 + 84274736720556630026311383931055307398820*I)/9671406556917033397649408,         (-5482217151670072904078758141270295025989 + 1818284338672191024475557065444481298568*I)/1208925819614629174706176,          (56564463395350195513805521309731217952281 - 360208541416798112109946262159695452898431*I)/19342813113834066795298816,        11*(1259539805728870739006416869463689438068 + 1409136581547898074455004171305324917387*I)/4835703278458516698824704,       5*(-123701190701414554945251071190688818343325 + 30997157322590424677294553832111902279712*I)/38685626227668133590597632,          (16130917381301373033736295883982414239781 - 32752041297570919727145380131926943374516*I)/9671406556917033397649408,          (650301385108223834347093740500375498354925 + 899526407681131828596801223402866051809258*I)/77371252455336267181195264],
#         [      (9011388245256140876590294262420614839483 + 8167917972423946282513000869327525382672*I)/1208925819614629174706176,       (-426393174084720190126376382194036323028924 + 180692224825757525982858693158209545430621*I)/9671406556917033397649408,     (24588556702197802674765733448108154175535 - 45091766022876486566421953254051868331066*I)/4835703278458516698824704,      (1872113939365285277373877183750416985089691 + 3030392393733212574744122057679633775773130*I)/77371252455336267181195264,    (-222173405538046189185754954524429864167549 - 75193157893478637039381059488387511299116*I)/19342813113834066795298816,        (2670821320766222522963689317316937579844558 - 2645837121493554383087981511645435472169191*I)/77371252455336267181195264,     5*(-2100110309556476773796963197283876204940 + 41957457246479840487980315496957337371937*I)/19342813113834066795298816,     (-5733743755499084165382383818991531258980593 - 3328949988392698205198574824396695027195732*I)/154742504910672534362390528,      (707827994365259025461378911159398206329247 - 265730616623227695108042528694302299777294*I)/77371252455336267181195264,    (-1442501604682933002895864804409322823788319 + 11504137805563265043376405214378288793343879*I)/309485009821345068724781056,         (-56130472299445561499538726459719629522285 - 61117552419727805035810982426639329818864*I)/9671406556917033397649408,    (39053692321126079849054272431599539429908717 - 10209127700342570953247177602860848130710666*I)/1237940039285380274899124224]])
    M = Matrix(S('''[
        [             -3/4,       45/32 - 37*I/16,         1/4 + I/2,      -129/64 - 9*I/64,      1/4 - 5*I/16,      65/128 + 87*I/64],
        [-149/64 + 49*I/32, -177/128 - 1369*I/128,  125/64 + 87*I/64, -2063/256 + 541*I/128,  85/256 - 33*I/16,  805/128 + 2415*I/512],
        [          1/2 - I,         9/4 + 55*I/16,              -3/4,       45/32 - 37*I/16,         1/4 + I/2,      -129/64 - 9*I/64],
        [   -5/8 - 39*I/16,   2473/256 + 137*I/64, -149/64 + 49*I/32, -177/128 - 1369*I/128,  125/64 + 87*I/64, -2063/256 + 541*I/128],
        [            1 + I,         -19/4 + 5*I/4,           1/2 - I,         9/4 + 55*I/16,              -3/4,       45/32 - 37*I/16],
        [         21/8 + I,    -537/64 + 143*I/16,    -5/8 - 39*I/16,   2473/256 + 137*I/64, -149/64 + 49*I/32, -177/128 - 1369*I/128]]'''))
    with dotprodsimp(True):
        assert M**10 == Matrix(S('''[
            [     7369525394972778926719607798014571861/604462909807314587353088 - 229284202061790301477392339912557559*I/151115727451828646838272,   -19704281515163975949388435612632058035/1208925819614629174706176 + 14319858347987648723768698170712102887*I/302231454903657293676544,      -3623281909451783042932142262164941211/604462909807314587353088 - 6039240602494288615094338643452320495*I/604462909807314587353088,    109260497799140408739847239685705357695/2417851639229258349412352 - 7427566006564572463236368211555511431*I/2417851639229258349412352, -16095803767674394244695716092817006641/2417851639229258349412352 + 10336681897356760057393429626719177583*I/1208925819614629174706176,    -42207883340488041844332828574359769743/2417851639229258349412352 - 182332262671671273188016400290188468499*I/4835703278458516698824704],
            [50566491050825573392726324995779608259/1208925819614629174706176 - 90047007594468146222002432884052362145*I/2417851639229258349412352,  74273703462900000967697427843983822011/1208925819614629174706176 + 265947522682943571171988741842776095421*I/1208925819614629174706176, -116900341394390200556829767923360888429/2417851639229258349412352 - 53153263356679268823910621474478756845*I/2417851639229258349412352, 195407378023867871243426523048612490249/1208925819614629174706176 - 1242417915995360200584837585002906728929*I/9671406556917033397649408,   -863597594389821970177319682495878193/302231454903657293676544 + 476936100741548328800725360758734300481*I/9671406556917033397649408, -3154451590535653853562472176601754835575/19342813113834066795298816 - 232909875490506237386836489998407329215*I/2417851639229258349412352],
            [   -1715444997702484578716037230949868543/302231454903657293676544 + 5009695651321306866158517287924120777*I/302231454903657293676544,     -30551582497996879620371947949342101301/604462909807314587353088 - 7632518367986526187139161303331519629*I/151115727451828646838272,           312680739924495153190604170938220575/18889465931478580854784 - 108664334509328818765959789219208459*I/75557863725914323419136,    -14693696966703036206178521686918865509/604462909807314587353088 + 72345386220900843930147151999899692401*I/1208925819614629174706176,  -8218872496728882299722894680635296519/1208925819614629174706176 - 16776782833358893712645864791807664983*I/1208925819614629174706176,      143237839169380078671242929143670635137/2417851639229258349412352 + 2883817094806115974748882735218469447*I/2417851639229258349412352],
            [   3087979417831061365023111800749855987/151115727451828646838272 + 34441942370802869368851419102423997089*I/604462909807314587353088, -148309181940158040917731426845476175667/604462909807314587353088 - 263987151804109387844966835369350904919*I/9671406556917033397649408,   50259518594816377378747711930008883165/1208925819614629174706176 - 95713974916869240305450001443767979653*I/2417851639229258349412352,  153466447023875527996457943521467271119/2417851639229258349412352 + 517285524891117105834922278517084871349*I/2417851639229258349412352,  -29184653615412989036678939366291205575/604462909807314587353088 - 27551322282526322041080173287022121083*I/1208925819614629174706176,   196404220110085511863671393922447671649/1208925819614629174706176 - 1204712019400186021982272049902206202145*I/9671406556917033397649408],
            [     -2632581805949645784625606590600098779/151115727451828646838272 - 589957435912868015140272627522612771*I/37778931862957161709568,     26727850893953715274702844733506310247/302231454903657293676544 - 10825791956782128799168209600694020481*I/302231454903657293676544,      -1036348763702366164044671908440791295/151115727451828646838272 + 3188624571414467767868303105288107375*I/151115727451828646838272,     -36814959939970644875593411585393242449/604462909807314587353088 - 18457555789119782404850043842902832647*I/302231454903657293676544,      12454491297984637815063964572803058647/604462909807314587353088 - 340489532842249733975074349495329171*I/302231454903657293676544,      -19547211751145597258386735573258916681/604462909807314587353088 + 87299583775782199663414539883938008933*I/1208925819614629174706176],
            [  -40281994229560039213253423262678393183/604462909807314587353088 - 2939986850065527327299273003299736641*I/604462909807314587353088, 331940684638052085845743020267462794181/2417851639229258349412352 - 284574901963624403933361315517248458969*I/1208925819614629174706176,      6453843623051745485064693628073010961/302231454903657293676544 + 36062454107479732681350914931391590957*I/604462909807314587353088,  -147665869053634695632880753646441962067/604462909807314587353088 - 305987938660447291246597544085345123927*I/9671406556917033397649408,  107821369195275772166593879711259469423/2417851639229258349412352 - 11645185518211204108659001435013326687*I/302231454903657293676544,     64121228424717666402009446088588091619/1208925819614629174706176 + 265557133337095047883844369272389762133*I/1208925819614629174706176]]'''))


def test_issue_17247_expression_blowup_5():
    M = Matrix(6, 6, lambda i, j: 1 + (-1)**(i+j)*I)
    with dotprodsimp(True):
        assert M.charpoly('x') == PurePoly(x**6 + (-6 - 6*I)*x**5 + 36*I*x**4, x, domain='EX')


def test_issue_17247_expression_blowup_6():
    M = Matrix(8, 8, [x+i for i in range (64)])
    with dotprodsimp(True):
        assert M.det('bareiss') == 0


def test_issue_17247_expression_blowup_7():
    M = Matrix(6, 6, lambda i, j: 1 + (-1)**(i+j)*I)
    with dotprodsimp(True):
        assert M.det('berkowitz') == 0


def test_issue_17247_expression_blowup_8():
    M = Matrix(8, 8, [x+i for i in range (64)])
    with dotprodsimp(True):
        assert M.det('lu') == 0


def test_issue_17247_expression_blowup_9():
    M = Matrix(8, 8, [x+i for i in range (64)])
    with dotprodsimp(True):
        assert M.rref() == (Matrix([
            [1, 0, -1, -2, -3, -4, -5, -6],
            [0, 1,  2,  3,  4,  5,  6,  7],
            [0, 0,  0,  0,  0,  0,  0,  0],
            [0, 0,  0,  0,  0,  0,  0,  0],
            [0, 0,  0,  0,  0,  0,  0,  0],
            [0, 0,  0,  0,  0,  0,  0,  0],
            [0, 0,  0,  0,  0,  0,  0,  0],
            [0, 0,  0,  0,  0,  0,  0,  0]]), (0, 1))


def test_issue_17247_expression_blowup_10():
    M = Matrix(6, 6, lambda i, j: 1 + (-1)**(i+j)*I)
    with dotprodsimp(True):
        assert M.cofactor(0, 0) == 0


def test_issue_17247_expression_blowup_11():
    M = Matrix(6, 6, lambda i, j: 1 + (-1)**(i+j)*I)
    with dotprodsimp(True):
        assert M.cofactor_matrix() == Matrix(6, 6, [0]*36)


def test_issue_17247_expression_blowup_12():
    M = Matrix(6, 6, lambda i, j: 1 + (-1)**(i+j)*I)
    with dotprodsimp(True):
        assert M.eigenvals() == {6: 1, 6*I: 1, 0: 4}


def test_issue_17247_expression_blowup_13():
    M = Matrix([
        [    0, 1 - x, x + 1, 1 - x],
        [1 - x, x + 1,     0, x + 1],
        [    0, 1 - x, x + 1, 1 - x],
        [    0,     0,     1 - x, 0]])

    ev = M.eigenvects()
    assert ev[0] == (0, 2, [Matrix([0, -1, 0, 1])])
    assert ev[1][0] == x - sqrt(2)*(x - 1) + 1
    assert ev[1][1] == 1
    assert ev[1][2][0].expand(deep=False, numer=True) == Matrix([
        [(-x + sqrt(2)*(x - 1) - 1)/(x - 1)],
        [-4*x/(x**2 - 2*x + 1) + (x + 1)*(x - sqrt(2)*(x - 1) + 1)/(x**2 - 2*x + 1)],
        [(-x + sqrt(2)*(x - 1) - 1)/(x - 1)],
        [1]
    ])

    assert ev[2][0] == x + sqrt(2)*(x - 1) + 1
    assert ev[2][1] == 1
    assert ev[2][2][0].expand(deep=False, numer=True) == Matrix([
        [(-x - sqrt(2)*(x - 1) - 1)/(x - 1)],
        [-4*x/(x**2 - 2*x + 1) + (x + 1)*(x + sqrt(2)*(x - 1) + 1)/(x**2 - 2*x + 1)],
        [(-x - sqrt(2)*(x - 1) - 1)/(x - 1)],
        [1]
    ])


def test_issue_17247_expression_blowup_14():
    M = Matrix(8, 8, ([1+x, 1-x]*4 + [1-x, 1+x]*4)*4)
    with dotprodsimp(True):
        assert M.echelon_form() == Matrix([
            [x + 1, 1 - x, x + 1, 1 - x, x + 1, 1 - x, x + 1, 1 - x],
            [    0,   4*x,     0,   4*x,     0,   4*x,     0,   4*x],
            [    0,     0,     0,     0,     0,     0,     0,     0],
            [    0,     0,     0,     0,     0,     0,     0,     0],
            [    0,     0,     0,     0,     0,     0,     0,     0],
            [    0,     0,     0,     0,     0,     0,     0,     0],
            [    0,     0,     0,     0,     0,     0,     0,     0],
            [    0,     0,     0,     0,     0,     0,     0,     0]])


def test_issue_17247_expression_blowup_15():
    M = Matrix(8, 8, ([1+x, 1-x]*4 + [1-x, 1+x]*4)*4)
    with dotprodsimp(True):
        assert M.rowspace() == [Matrix([[x + 1, 1 - x, x + 1, 1 - x, x + 1, 1 - x, x + 1, 1 - x]]), Matrix([[0, 4*x, 0, 4*x, 0, 4*x, 0, 4*x]])]


def test_issue_17247_expression_blowup_16():
    M = Matrix(8, 8, ([1+x, 1-x]*4 + [1-x, 1+x]*4)*4)
    with dotprodsimp(True):
        assert M.columnspace() == [Matrix([[x + 1],[1 - x],[x + 1],[1 - x],[x + 1],[1 - x],[x + 1],[1 - x]]), Matrix([[1 - x],[x + 1],[1 - x],[x + 1],[1 - x],[x + 1],[1 - x],[x + 1]])]


def test_issue_17247_expression_blowup_17():
    M = Matrix(8, 8, [x+i for i in range (64)])
    with dotprodsimp(True):
        assert M.nullspace() == [
            Matrix([[1],[-2],[1],[0],[0],[0],[0],[0]]),
            Matrix([[2],[-3],[0],[1],[0],[0],[0],[0]]),
            Matrix([[3],[-4],[0],[0],[1],[0],[0],[0]]),
            Matrix([[4],[-5],[0],[0],[0],[1],[0],[0]]),
            Matrix([[5],[-6],[0],[0],[0],[0],[1],[0]]),
            Matrix([[6],[-7],[0],[0],[0],[0],[0],[1]])]


def test_issue_17247_expression_blowup_18():
    M = Matrix(6, 6, ([1+x, 1-x]*3 + [1-x, 1+x]*3)*3)
    with dotprodsimp(True):
        assert not M.is_nilpotent()


def test_issue_17247_expression_blowup_19():
    M = Matrix(S('''[
        [             -3/4,                     0,         1/4 + I/2,                     0],
        [                0, -177/128 - 1369*I/128,                 0, -2063/256 + 541*I/128],
        [          1/2 - I,                     0,                 0,                     0],
        [                0,                     0,                 0, -177/128 - 1369*I/128]]'''))
    with dotprodsimp(True):
        assert not M.is_diagonalizable()


def test_issue_17247_expression_blowup_20():
    M = Matrix([
    [x + 1,  1 - x,      0,      0],
    [1 - x,  x + 1,      0,  x + 1],
    [    0,  1 - x,  x + 1,      0],
    [    0,      0,      0,  x + 1]])
    with dotprodsimp(True):
        assert M.diagonalize() == (Matrix([
            [1,  1, 0, (x + 1)/(x - 1)],
            [1, -1, 0,               0],
            [1,  1, 1,               0],
            [0,  0, 0,               1]]),
            Matrix([
            [2,   0,     0,     0],
            [0, 2*x,     0,     0],
            [0,   0, x + 1,     0],
            [0,   0,     0, x + 1]]))


def test_issue_17247_expression_blowup_21():
    M = Matrix(S('''[
        [             -3/4,       45/32 - 37*I/16,                   0,                     0],
        [-149/64 + 49*I/32, -177/128 - 1369*I/128,                   0, -2063/256 + 541*I/128],
        [                0,         9/4 + 55*I/16, 2473/256 + 137*I/64,                     0],
        [                0,                     0,                   0, -177/128 - 1369*I/128]]'''))
    with dotprodsimp(True):
        assert M.inv(method='GE') == Matrix(S('''[
            [-26194832/3470993 - 31733264*I/3470993, 156352/3470993 + 10325632*I/3470993, 0, -7741283181072/3306971225785 + 2999007604624*I/3306971225785],
            [4408224/3470993 - 9675328*I/3470993, -2422272/3470993 + 1523712*I/3470993, 0, -1824666489984/3306971225785 - 1401091949952*I/3306971225785],
            [-26406945676288/22270005630769 + 10245925485056*I/22270005630769, 7453523312640/22270005630769 + 1601616519168*I/22270005630769, 633088/6416033 - 140288*I/6416033, 872209227109521408/21217636514687010905 + 6066405081802389504*I/21217636514687010905],
            [0, 0, 0, -11328/952745 + 87616*I/952745]]'''))


def test_issue_17247_expression_blowup_22():
    M = Matrix(S('''[
        [             -3/4,       45/32 - 37*I/16,                   0,                     0],
        [-149/64 + 49*I/32, -177/128 - 1369*I/128,                   0, -2063/256 + 541*I/128],
        [                0,         9/4 + 55*I/16, 2473/256 + 137*I/64,                     0],
        [                0,                     0,                   0, -177/128 - 1369*I/128]]'''))
    with dotprodsimp(True):
        assert M.inv(method='LU') == Matrix(S('''[
            [-26194832/3470993 - 31733264*I/3470993, 156352/3470993 + 10325632*I/3470993, 0, -7741283181072/3306971225785 + 2999007604624*I/3306971225785],
            [4408224/3470993 - 9675328*I/3470993, -2422272/3470993 + 1523712*I/3470993, 0, -1824666489984/3306971225785 - 1401091949952*I/3306971225785],
            [-26406945676288/22270005630769 + 10245925485056*I/22270005630769, 7453523312640/22270005630769 + 1601616519168*I/22270005630769, 633088/6416033 - 140288*I/6416033, 872209227109521408/21217636514687010905 + 6066405081802389504*I/21217636514687010905],
            [0, 0, 0, -11328/952745 + 87616*I/952745]]'''))


def test_issue_17247_expression_blowup_23():
    M = Matrix(S('''[
        [             -3/4,       45/32 - 37*I/16,                   0,                     0],
        [-149/64 + 49*I/32, -177/128 - 1369*I/128,                   0, -2063/256 + 541*I/128],
        [                0,         9/4 + 55*I/16, 2473/256 + 137*I/64,                     0],
        [                0,                     0,                   0, -177/128 - 1369*I/128]]'''))
    with dotprodsimp(True):
        assert M.inv(method='ADJ').expand() == Matrix(S('''[
            [-26194832/3470993 - 31733264*I/3470993, 156352/3470993 + 10325632*I/3470993, 0, -7741283181072/3306971225785 + 2999007604624*I/3306971225785],
            [4408224/3470993 - 9675328*I/3470993, -2422272/3470993 + 1523712*I/3470993, 0, -1824666489984/3306971225785 - 1401091949952*I/3306971225785],
            [-26406945676288/22270005630769 + 10245925485056*I/22270005630769, 7453523312640/22270005630769 + 1601616519168*I/22270005630769, 633088/6416033 - 140288*I/6416033, 872209227109521408/21217636514687010905 + 6066405081802389504*I/21217636514687010905],
            [0, 0, 0, -11328/952745 + 87616*I/952745]]'''))


def test_issue_17247_expression_blowup_24():
    M = SparseMatrix(S('''[
        [             -3/4,       45/32 - 37*I/16,                   0,                     0],
        [-149/64 + 49*I/32, -177/128 - 1369*I/128,                   0, -2063/256 + 541*I/128],
        [                0,         9/4 + 55*I/16, 2473/256 + 137*I/64,                     0],
        [                0,                     0,                   0, -177/128 - 1369*I/128]]'''))
    with dotprodsimp(True):
        assert M.inv(method='CH') == Matrix(S('''[
            [-26194832/3470993 - 31733264*I/3470993, 156352/3470993 + 10325632*I/3470993, 0, -7741283181072/3306971225785 + 2999007604624*I/3306971225785],
            [4408224/3470993 - 9675328*I/3470993, -2422272/3470993 + 1523712*I/3470993, 0, -1824666489984/3306971225785 - 1401091949952*I/3306971225785],
            [-26406945676288/22270005630769 + 10245925485056*I/22270005630769, 7453523312640/22270005630769 + 1601616519168*I/22270005630769, 633088/6416033 - 140288*I/6416033, 872209227109521408/21217636514687010905 + 6066405081802389504*I/21217636514687010905],
            [0, 0, 0, -11328/952745 + 87616*I/952745]]'''))


def test_issue_17247_expression_blowup_25():
    M = SparseMatrix(S('''[
        [             -3/4,       45/32 - 37*I/16,                   0,                     0],
        [-149/64 + 49*I/32, -177/128 - 1369*I/128,                   0, -2063/256 + 541*I/128],
        [                0,         9/4 + 55*I/16, 2473/256 + 137*I/64,                     0],
        [                0,                     0,                   0, -177/128 - 1369*I/128]]'''))
    with dotprodsimp(True):
        assert M.inv(method='LDL') == Matrix(S('''[
            [-26194832/3470993 - 31733264*I/3470993, 156352/3470993 + 10325632*I/3470993, 0, -7741283181072/3306971225785 + 2999007604624*I/3306971225785],
            [4408224/3470993 - 9675328*I/3470993, -2422272/3470993 + 1523712*I/3470993, 0, -1824666489984/3306971225785 - 1401091949952*I/3306971225785],
            [-26406945676288/22270005630769 + 10245925485056*I/22270005630769, 7453523312640/22270005630769 + 1601616519168*I/22270005630769, 633088/6416033 - 140288*I/6416033, 872209227109521408/21217636514687010905 + 6066405081802389504*I/21217636514687010905],
            [0, 0, 0, -11328/952745 + 87616*I/952745]]'''))


def test_issue_17247_expression_blowup_26():
    M = Matrix(S('''[
        [             -3/4,       45/32 - 37*I/16,         1/4 + I/2,      -129/64 - 9*I/64,      1/4 - 5*I/16,      65/128 + 87*I/64,         -9/32 - I/16,      183/256 - 97*I/128],
        [-149/64 + 49*I/32, -177/128 - 1369*I/128,  125/64 + 87*I/64, -2063/256 + 541*I/128,  85/256 - 33*I/16,  805/128 + 2415*I/512, -219/128 + 115*I/256, 6301/4096 - 6609*I/1024],
        [          1/2 - I,         9/4 + 55*I/16,              -3/4,       45/32 - 37*I/16,         1/4 + I/2,      -129/64 - 9*I/64,         1/4 - 5*I/16,        65/128 + 87*I/64],
        [   -5/8 - 39*I/16,   2473/256 + 137*I/64, -149/64 + 49*I/32, -177/128 - 1369*I/128,  125/64 + 87*I/64, -2063/256 + 541*I/128,     85/256 - 33*I/16,    805/128 + 2415*I/512],
        [            1 + I,         -19/4 + 5*I/4,           1/2 - I,         9/4 + 55*I/16,              -3/4,       45/32 - 37*I/16,            1/4 + I/2,        -129/64 - 9*I/64],
        [         21/8 + I,    -537/64 + 143*I/16,    -5/8 - 39*I/16,   2473/256 + 137*I/64, -149/64 + 49*I/32, -177/128 - 1369*I/128,     125/64 + 87*I/64,   -2063/256 + 541*I/128],
        [               -2,         17/4 - 13*I/2,             1 + I,         -19/4 + 5*I/4,           1/2 - I,         9/4 + 55*I/16,                 -3/4,         45/32 - 37*I/16],
        [     1/4 + 13*I/4,    -825/64 - 147*I/32,          21/8 + I,    -537/64 + 143*I/16,    -5/8 - 39*I/16,   2473/256 + 137*I/64,    -149/64 + 49*I/32,   -177/128 - 1369*I/128]]'''))
    with dotprodsimp(True):
        assert M.rank() == 4


def test_issue_17247_expression_blowup_27():
    M = Matrix([
        [    0, 1 - x, x + 1, 1 - x],
        [1 - x, x + 1,     0, x + 1],
        [    0, 1 - x, x + 1, 1 - x],
        [    0,     0,     1 - x, 0]])
    with dotprodsimp(True):
        P, J = M.jordan_form()
        assert P.expand() == Matrix(S('''[
            [    0,  4*x/(x**2 - 2*x + 1), -(-17*x**4 + 12*sqrt(2)*x**4 - 4*sqrt(2)*x**3 + 6*x**3 - 6*x - 4*sqrt(2)*x + 12*sqrt(2) + 17)/(-7*x**4 + 5*sqrt(2)*x**4 - 6*sqrt(2)*x**3 + 8*x**3 - 2*x**2 + 8*x + 6*sqrt(2)*x - 5*sqrt(2) - 7), -(12*sqrt(2)*x**4 + 17*x**4 - 6*x**3 - 4*sqrt(2)*x**3 - 4*sqrt(2)*x + 6*x - 17 + 12*sqrt(2))/(7*x**4 + 5*sqrt(2)*x**4 - 6*sqrt(2)*x**3 - 8*x**3 + 2*x**2 - 8*x + 6*sqrt(2)*x - 5*sqrt(2) + 7)],
            [x - 1, x/(x - 1) + 1/(x - 1),                       (-7*x**3 + 5*sqrt(2)*x**3 - x**2 + sqrt(2)*x**2 - sqrt(2)*x - x - 5*sqrt(2) - 7)/(-3*x**3 + 2*sqrt(2)*x**3 - 2*sqrt(2)*x**2 + 3*x**2 + 2*sqrt(2)*x + 3*x - 3 - 2*sqrt(2)),                       (7*x**3 + 5*sqrt(2)*x**3 + x**2 + sqrt(2)*x**2 - sqrt(2)*x + x - 5*sqrt(2) + 7)/(2*sqrt(2)*x**3 + 3*x**3 - 3*x**2 - 2*sqrt(2)*x**2 - 3*x + 2*sqrt(2)*x - 2*sqrt(2) + 3)],
            [    0,                     1,                                                                                            -(-3*x**2 + 2*sqrt(2)*x**2 + 2*x - 3 - 2*sqrt(2))/(-x**2 + sqrt(2)*x**2 - 2*sqrt(2)*x + 1 + sqrt(2)),                                                                                            -(2*sqrt(2)*x**2 + 3*x**2 - 2*x - 2*sqrt(2) + 3)/(x**2 + sqrt(2)*x**2 - 2*sqrt(2)*x - 1 + sqrt(2))],
            [1 - x,                     0,                                                                                                                                                                                               1,                                                                                                                                                                                             1]]''')).expand()
        assert J == Matrix(S('''[
            [0, 1,                       0,                       0],
            [0, 0,                       0,                       0],
            [0, 0, x - sqrt(2)*(x - 1) + 1,                       0],
            [0, 0,                       0, x + sqrt(2)*(x - 1) + 1]]'''))


def test_issue_17247_expression_blowup_28():
    M = Matrix(S('''[
        [             -3/4,       45/32 - 37*I/16,                   0,                     0],
        [-149/64 + 49*I/32, -177/128 - 1369*I/128,                   0, -2063/256 + 541*I/128],
        [                0,         9/4 + 55*I/16, 2473/256 + 137*I/64,                     0],
        [                0,                     0,                   0, -177/128 - 1369*I/128]]'''))
    with dotprodsimp(True):
        assert M.singular_values() == S('''[
            sqrt(14609315/131072 + sqrt(64789115132571/2147483648 - 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3) + 76627253330829751075/(35184372088832*sqrt(64789115132571/4294967296 + 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)) + 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3))) - 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)))/2 + sqrt(64789115132571/4294967296 + 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)) + 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3))/2),
            sqrt(14609315/131072 - sqrt(64789115132571/2147483648 - 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3) + 76627253330829751075/(35184372088832*sqrt(64789115132571/4294967296 + 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)) + 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3))) - 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)))/2 + sqrt(64789115132571/4294967296 + 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)) + 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3))/2),
            sqrt(14609315/131072 - sqrt(64789115132571/4294967296 + 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)) + 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3))/2 + sqrt(64789115132571/2147483648 - 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3) - 76627253330829751075/(35184372088832*sqrt(64789115132571/4294967296 + 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)) + 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3))) - 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)))/2),
            sqrt(14609315/131072 - sqrt(64789115132571/4294967296 + 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)) + 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3))/2 - sqrt(64789115132571/2147483648 - 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3) - 76627253330829751075/(35184372088832*sqrt(64789115132571/4294967296 + 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)) + 2*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3))) - 3546944054712886603889144627/(110680464442257309696*(25895222463957462655758224991455280215303/633825300114114700748351602688 + sqrt(1213909058710955930446995195883114969038524625997915131236390724543989220134670)*I/22282920707136844948184236032)**(1/3)))/2)]''')


def test_issue_16823():
    # This still needs to be fixed if not using dotprodsimp.
    M = Matrix(S('''[
        [1+I,-19/4+5/4*I,1/2-I,9/4+55/16*I,-3/4,45/32-37/16*I,1/4+1/2*I,-129/64-9/64*I,1/4-5/16*I,65/128+87/64*I,-9/32-1/16*I,183/256-97/128*I,3/64+13/64*I,-23/32-59/256*I,15/128-3/32*I,19/256+551/1024*I],
        [21/8+I,-537/64+143/16*I,-5/8-39/16*I,2473/256+137/64*I,-149/64+49/32*I,-177/128-1369/128*I,125/64+87/64*I,-2063/256+541/128*I,85/256-33/16*I,805/128+2415/512*I,-219/128+115/256*I,6301/4096-6609/1024*I,119/128+143/128*I,-10879/2048+4343/4096*I,129/256-549/512*I,42533/16384+29103/8192*I],
        [-2,17/4-13/2*I,1+I,-19/4+5/4*I,1/2-I,9/4+55/16*I,-3/4,45/32-37/16*I,1/4+1/2*I,-129/64-9/64*I,1/4-5/16*I,65/128+87/64*I,-9/32-1/16*I,183/256-97/128*I,3/64+13/64*I,-23/32-59/256*I],
        [1/4+13/4*I,-825/64-147/32*I,21/8+I,-537/64+143/16*I,-5/8-39/16*I,2473/256+137/64*I,-149/64+49/32*I,-177/128-1369/128*I,125/64+87/64*I,-2063/256+541/128*I,85/256-33/16*I,805/128+2415/512*I,-219/128+115/256*I,6301/4096-6609/1024*I,119/128+143/128*I,-10879/2048+4343/4096*I],
        [-4*I,27/2+6*I,-2,17/4-13/2*I,1+I,-19/4+5/4*I,1/2-I,9/4+55/16*I,-3/4,45/32-37/16*I,1/4+1/2*I,-129/64-9/64*I,1/4-5/16*I,65/128+87/64*I,-9/32-1/16*I,183/256-97/128*I],
        [1/4+5/2*I,-23/8-57/16*I,1/4+13/4*I,-825/64-147/32*I,21/8+I,-537/64+143/16*I,-5/8-39/16*I,2473/256+137/64*I,-149/64+49/32*I,-177/128-1369/128*I,125/64+87/64*I,-2063/256+541/128*I,85/256-33/16*I,805/128+2415/512*I,-219/128+115/256*I,6301/4096-6609/1024*I],
        [-4,9-5*I,-4*I,27/2+6*I,-2,17/4-13/2*I,1+I,-19/4+5/4*I,1/2-I,9/4+55/16*I,-3/4,45/32-37/16*I,1/4+1/2*I,-129/64-9/64*I,1/4-5/16*I,65/128+87/64*I],
        [-2*I,119/8+29/4*I,1/4+5/2*I,-23/8-57/16*I,1/4+13/4*I,-825/64-147/32*I,21/8+I,-537/64+143/16*I,-5/8-39/16*I,2473/256+137/64*I,-149/64+49/32*I,-177/128-1369/128*I,125/64+87/64*I,-2063/256+541/128*I,85/256-33/16*I,805/128+2415/512*I],
        [0,-6,-4,9-5*I,-4*I,27/2+6*I,-2,17/4-13/2*I,1+I,-19/4+5/4*I,1/2-I,9/4+55/16*I,-3/4,45/32-37/16*I,1/4+1/2*I,-129/64-9/64*I],
        [1,-9/4+3*I,-2*I,119/8+29/4*I,1/4+5/2*I,-23/8-57/16*I,1/4+13/4*I,-825/64-147/32*I,21/8+I,-537/64+143/16*I,-5/8-39/16*I,2473/256+137/64*I,-149/64+49/32*I,-177/128-1369/128*I,125/64+87/64*I,-2063/256+541/128*I],
        [0,-4*I,0,-6,-4,9-5*I,-4*I,27/2+6*I,-2,17/4-13/2*I,1+I,-19/4+5/4*I,1/2-I,9/4+55/16*I,-3/4,45/32-37/16*I],
        [0,1/4+1/2*I,1,-9/4+3*I,-2*I,119/8+29/4*I,1/4+5/2*I,-23/8-57/16*I,1/4+13/4*I,-825/64-147/32*I,21/8+I,-537/64+143/16*I,-5/8-39/16*I,2473/256+137/64*I,-149/64+49/32*I,-177/128-1369/128*I]]'''))
    with dotprodsimp(True):
        assert M.rank() == 8


def test_issue_18531():
    # solve_linear_system still needs fixing but the rref works.
    M = Matrix([
        [1, 1, 1, 1, 1, 0, 1, 0, 0],
        [1 + sqrt(2), -1 + sqrt(2), 1 - sqrt(2), -sqrt(2) - 1, 1, 1, -1, 1, 1],
        [-5 + 2*sqrt(2), -5 - 2*sqrt(2), -5 - 2*sqrt(2), -5 + 2*sqrt(2), -7, 2, -7, -2, 0],
        [-3*sqrt(2) - 1, 1 - 3*sqrt(2), -1 + 3*sqrt(2), 1 + 3*sqrt(2), -7, -5, 7, -5, 3],
        [7 - 4*sqrt(2), 4*sqrt(2) + 7, 4*sqrt(2) + 7, 7 - 4*sqrt(2), 7, -12, 7, 12, 0],
        [-1 + 3*sqrt(2), 1 + 3*sqrt(2), -3*sqrt(2) - 1, 1 - 3*sqrt(2), 7, -5, -7, -5, 3],
        [-3 + 2*sqrt(2), -3 - 2*sqrt(2), -3 - 2*sqrt(2), -3 + 2*sqrt(2), -1, 2, -1, -2, 0],
        [1 - sqrt(2), -sqrt(2) - 1, 1 + sqrt(2), -1 + sqrt(2), -1, 1, 1, 1, 1]
        ])
    with dotprodsimp(True):
        assert M.rref() == (Matrix([
            [1, 0, 0, 0, 0, 0, 0, 0,  S(1)/2],
            [0, 1, 0, 0, 0, 0, 0, 0, -S(1)/2],
            [0, 0, 1, 0, 0, 0, 0, 0,  S(1)/2],
            [0, 0, 0, 1, 0, 0, 0, 0, -S(1)/2],
            [0, 0, 0, 0, 1, 0, 0, 0,    0],
            [0, 0, 0, 0, 0, 1, 0, 0, -S(1)/2],
            [0, 0, 0, 0, 0, 0, 1, 0,    0],
            [0, 0, 0, 0, 0, 0, 0, 1, -S(1)/2]]), (0, 1, 2, 3, 4, 5, 6, 7))


def test_creation():
    raises(ValueError, lambda: Matrix(5, 5, range(20)))
    raises(ValueError, lambda: Matrix(5, -1, []))
    raises(IndexError, lambda: Matrix((1, 2))[2])
    with raises(IndexError):
        Matrix((1, 2))[3] = 5

    assert Matrix() == Matrix([]) == Matrix(0, 0, [])
    assert Matrix([[]]) == Matrix(1, 0, [])
    assert Matrix([[], []]) == Matrix(2, 0, [])

    # anything used to be allowed in a matrix
    with warns_deprecated_sympy():
        assert Matrix([[[1], (2,)]]).tolist() == [[[1], (2,)]]
    with warns_deprecated_sympy():
        assert Matrix([[[1], (2,)]]).T.tolist() == [[[1]], [(2,)]]
    M = Matrix([[0]])
    with warns_deprecated_sympy():
        M[0, 0] = S.EmptySet

    a = Matrix([[x, 0], [0, 0]])
    m = a
    assert m.cols == m.rows
    assert m.cols == 2
    assert m[:] == [x, 0, 0, 0]

    b = Matrix(2, 2, [x, 0, 0, 0])
    m = b
    assert m.cols == m.rows
    assert m.cols == 2
    assert m[:] == [x, 0, 0, 0]

    assert a == b

    assert Matrix(b) == b

    c23 = Matrix(2, 3, range(1, 7))
    c13 = Matrix(1, 3, range(7, 10))
    c = Matrix([c23, c13])
    assert c.cols == 3
    assert c.rows == 3
    assert c[:] == [1, 2, 3, 4, 5, 6, 7, 8, 9]

    assert Matrix(eye(2)) == eye(2)
    assert ImmutableMatrix(ImmutableMatrix(eye(2))) == ImmutableMatrix(eye(2))
    assert ImmutableMatrix(c) == c.as_immutable()
    assert Matrix(ImmutableMatrix(c)) == ImmutableMatrix(c).as_mutable()

    assert c is not Matrix(c)

    dat = [[ones(3,2), ones(3,3)*2], [ones(2,3)*3, ones(2,2)*4]]
    M = Matrix(dat)
    assert M == Matrix([
        [1, 1, 2, 2, 2],
        [1, 1, 2, 2, 2],
        [1, 1, 2, 2, 2],
        [3, 3, 3, 4, 4],
        [3, 3, 3, 4, 4]])
    assert M.tolist() != dat
    # keep block form if evaluate=False
    assert Matrix(dat, evaluate=False).tolist() == dat
    A = MatrixSymbol("A", 2, 2)
    dat = [ones(2), A]
    assert Matrix(dat) == Matrix([
    [      1,       1],
    [      1,       1],
    [A[0, 0], A[0, 1]],
    [A[1, 0], A[1, 1]]])
    with warns_deprecated_sympy():
        assert Matrix(dat, evaluate=False).tolist() == [[i] for i in dat]

    # 0-dim tolerance
    assert Matrix([ones(2), ones(0)]) == Matrix([ones(2)])
    raises(ValueError, lambda: Matrix([ones(2), ones(0, 3)]))
    raises(ValueError, lambda: Matrix([ones(2), ones(3, 0)]))

    # mix of Matrix and iterable
    M = Matrix([[1, 2], [3, 4]])
    M2 = Matrix([M, (5, 6)])
    assert M2 == Matrix([[1, 2], [3, 4], [5, 6]])


def test_irregular_block():
    assert Matrix.irregular(3, ones(2,1), ones(3,3)*2, ones(2,2)*3,
        ones(1,1)*4, ones(2,2)*5, ones(1,2)*6, ones(1,2)*7) == Matrix([
        [1, 2, 2, 2, 3, 3],
        [1, 2, 2, 2, 3, 3],
        [4, 2, 2, 2, 5, 5],
        [6, 6, 7, 7, 5, 5]])


def test_slicing():
    m0 = eye(4)
    assert m0[:3, :3] == eye(3)
    assert m0[2:4, 0:2] == zeros(2)

    m1 = Matrix(3, 3, lambda i, j: i + j)
    assert m1[0, :] == Matrix(1, 3, (0, 1, 2))
    assert m1[1:3, 1] == Matrix(2, 1, (2, 3))

    m2 = Matrix([[0, 1, 2, 3], [4, 5, 6, 7], [8, 9, 10, 11], [12, 13, 14, 15]])
    assert m2[:, -1] == Matrix(4, 1, [3, 7, 11, 15])
    assert m2[-2:, :] == Matrix([[8, 9, 10, 11], [12, 13, 14, 15]])


def test_submatrix_assignment():
    m = zeros(4)
    m[2:4, 2:4] = eye(2)
    assert m == Matrix(((0, 0, 0, 0),
                        (0, 0, 0, 0),
                        (0, 0, 1, 0),
                        (0, 0, 0, 1)))
    m[:2, :2] = eye(2)
    assert m == eye(4)
    m[:, 0] = Matrix(4, 1, (1, 2, 3, 4))
    assert m == Matrix(((1, 0, 0, 0),
                        (2, 1, 0, 0),
                        (3, 0, 1, 0),
                        (4, 0, 0, 1)))
    m[:, :] = zeros(4)
    assert m == zeros(4)
    m[:, :] = [(1, 2, 3, 4), (5, 6, 7, 8), (9, 10, 11, 12), (13, 14, 15, 16)]
    assert m == Matrix(((1, 2, 3, 4),
                        (5, 6, 7, 8),
                        (9, 10, 11, 12),
                        (13, 14, 15, 16)))
    m[:2, 0] = [0, 0]
    assert m == Matrix(((0, 2, 3, 4),
                        (0, 6, 7, 8),
                        (9, 10, 11, 12),
                        (13, 14, 15, 16)))


def test_reshape():
    m0 = eye(3)
    assert m0.reshape(1, 9) == Matrix(1, 9, (1, 0, 0, 0, 1, 0, 0, 0, 1))
    m1 = Matrix(3, 4, lambda i, j: i + j)
    assert m1.reshape(
        4, 3) == Matrix(((0, 1, 2), (3, 1, 2), (3, 4, 2), (3, 4, 5)))
    assert m1.reshape(2, 6) == Matrix(((0, 1, 2, 3, 1, 2), (3, 4, 2, 3, 4, 5)))


def test_applyfunc():
    m0 = eye(3)
    assert m0.applyfunc(lambda x: 2*x) == eye(3)*2
    assert m0.applyfunc(lambda x: 0) == zeros(3)


def test_expand():
    m0 = Matrix([[x*(x + y), 2], [((x + y)*y)*x, x*(y + x*(x + y))]])
    # Test if expand() returns a matrix
    m1 = m0.expand()
    assert m1 == Matrix(
        [[x*y + x**2, 2], [x*y**2 + y*x**2, x*y + y*x**2 + x**3]])

    a = Symbol('a', real=True)

    assert Matrix([exp(I*a)]).expand(complex=True) == \
        Matrix([cos(a) + I*sin(a)])

    assert Matrix([[0, 1, 2], [0, 0, -1], [0, 0, 0]]).exp() == Matrix([
        [1, 1, Rational(3, 2)],
        [0, 1, -1],
        [0, 0, 1]]
    )


def test_refine():
    m0 = Matrix([[Abs(x)**2, sqrt(x**2)],
                [sqrt(x**2)*Abs(y)**2, sqrt(y**2)*Abs(x)**2]])
    m1 = m0.refine(Q.real(x) & Q.real(y))
    assert m1 == Matrix([[x**2, Abs(x)], [y**2*Abs(x), x**2*Abs(y)]])

    m1 = m0.refine(Q.positive(x) & Q.positive(y))
    assert m1 == Matrix([[x**2, x], [x*y**2, x**2*y]])

    m1 = m0.refine(Q.negative(x) & Q.negative(y))
    assert m1 == Matrix([[x**2, -x], [-x*y**2, -x**2*y]])


def test_random():
    M = randMatrix(3, 3)
    M = randMatrix(3, 3, seed=3)
    assert M == randMatrix(3, 3, seed=3)

    M = randMatrix(3, 4, 0, 150)
    M = randMatrix(3, seed=4, symmetric=True)
    assert M == randMatrix(3, seed=4, symmetric=True)

    S = M.copy()
    S.simplify()
    assert S == M  # doesn't fail when elements are Numbers, not int

    rng = random.Random(4)
    assert M == randMatrix(3, symmetric=True, prng=rng)

    # Ensure symmetry
    for size in (10, 11): # Test odd and even
        for percent in (100, 70, 30):
            M = randMatrix(size, symmetric=True, percent=percent, prng=rng)
            assert M == M.T

    M = randMatrix(10, min=1, percent=70)
    zero_count = 0
    for i in range(M.shape[0]):
        for j in range(M.shape[1]):
            if M[i, j] == 0:
                zero_count += 1
    assert zero_count == 30


def test_inverse():
    A = eye(4)
    assert A.inv() == eye(4)
    assert A.inv(method="LU") == eye(4)
    assert A.inv(method="ADJ") == eye(4)
    assert A.inv(method="CH") == eye(4)
    assert A.inv(method="LDL") == eye(4)
    assert A.inv(method="QR") == eye(4)
    A = Matrix([[2, 3, 5],
                [3, 6, 2],
                [8, 3, 6]])
    Ainv = A.inv()
    assert A*Ainv == eye(3)
    assert A.inv(method="LU") == Ainv
    assert A.inv(method="ADJ") == Ainv
    assert A.inv(method="CH") == Ainv
    assert A.inv(method="LDL") == Ainv
    assert A.inv(method="QR") == Ainv

    AA = Matrix([[0, 0, 0, 0, 1, 0, 0, 1, 0, 1, 1, 1, 0, 0, 0, 1, 1, 0, 0, 0, 0, 1, 1, 0, 0],
            [1, 0, 1, 0, 0, 1, 0, 0, 1, 0, 1, 0, 1, 1, 1, 0, 1, 1, 0, 0, 0, 0, 0, 1, 0],
            [1, 1, 1, 1, 0, 1, 0, 0, 1, 1, 0, 0, 0, 1, 1, 1, 1, 0, 0, 1, 1, 0, 0, 1, 1],
            [1, 0, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0],
            [1, 0, 0, 1, 1, 1, 0, 1, 0, 0, 1, 0, 0, 1, 1, 1, 0, 1, 0, 0, 0, 1, 0, 0, 0],
            [1, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 0, 1, 0, 1, 1, 1, 1, 0, 0, 0, 1, 1, 0, 1],
            [0, 1, 1, 1, 0, 0, 0, 0, 1, 1, 1, 0, 1, 0, 1, 0, 1, 0, 1, 0, 1, 0, 0, 1, 0],
            [1, 1, 0, 0, 0, 1, 1, 0, 0, 1, 1, 0, 1, 1, 0, 1, 1, 0, 1, 0, 1, 1, 0, 1, 1],
            [0, 0, 1, 1, 1, 0, 0, 1, 1, 0, 1, 1, 0, 1, 0, 0, 0, 0, 0, 0, 1, 0, 1, 0, 1],
            [1, 0, 1, 1, 1, 0, 1, 0, 0, 1, 1, 0, 1, 1, 0, 1, 1, 1, 1, 0, 1, 1, 1, 0, 0],
            [0, 1, 1, 0, 0, 0, 0, 1, 1, 0, 1, 0, 1, 0, 1, 1, 0, 0, 1, 0, 0, 0, 1, 1, 0],
            [1, 0, 0, 1, 0, 1, 0, 1, 1, 0, 0, 1, 0, 1, 0, 1, 0, 1, 0, 1, 0, 1, 1, 0, 0],
            [0, 1, 1, 0, 0, 1, 1, 0, 0, 1, 1, 0, 1, 0, 0, 0, 1, 0, 0, 1, 0, 1, 1, 0, 1],
            [1, 1, 1, 0, 1, 0, 0, 1, 0, 0, 1, 0, 1, 1, 0, 0, 1, 1, 1, 0, 1, 0, 1, 1, 0],
            [0, 1, 0, 0, 0, 0, 1, 1, 0, 1, 1, 0, 0, 0, 0, 1, 0, 1, 1, 1, 0, 0, 1, 0, 0],
            [1, 1, 1, 1, 0, 1, 1, 1, 1, 1, 1, 0, 0, 0, 1, 0, 1, 0, 1, 0, 0, 1, 0, 0, 0],
            [0, 0, 0, 1, 0, 1, 1, 0, 1, 0, 0, 0, 0, 0, 1, 0, 1, 1, 0, 1, 0, 1, 1, 0, 1],
            [0, 1, 0, 1, 0, 0, 0, 0, 1, 1, 1, 0, 0, 1, 1, 1, 0, 0, 1, 0, 1, 1, 0, 1, 1],
            [1, 0, 1, 1, 0, 1, 0, 1, 1, 0, 1, 1, 1, 0, 1, 0, 0, 0, 0, 1, 0, 0, 1, 0, 1],
            [0, 0, 0, 1, 0, 0, 1, 1, 0, 1, 1, 0, 1, 0, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0],
            [1, 0, 0, 0, 0, 1, 0, 1, 0, 0, 1, 1, 0, 1, 0, 1, 1, 1, 1, 0, 1, 1, 1, 1, 1],
            [0, 0, 1, 1, 0, 1, 0, 0, 0, 1, 0, 0, 1, 0, 0, 1, 0, 0, 0, 1, 0, 1, 0, 1, 1],
            [0, 0, 0, 0, 0, 1, 1, 1, 0, 1, 1, 1, 0, 0, 0, 1, 0, 1, 1, 0, 1, 1, 0, 0, 0],
            [0, 0, 1, 0, 1, 1, 0, 1, 1, 1, 0, 0, 1, 0, 1, 0, 0, 1, 0, 1, 0, 0, 0, 0, 0],
            [0, 0, 0, 1, 1, 0, 1, 1, 1, 0, 0, 1, 0, 1, 0, 1, 1, 0, 1, 1, 1, 0, 0, 1, 0]])
    assert AA.inv(method="BLOCK") * AA == eye(AA.shape[0])
    # test that immutability is not a problem
    cls = ImmutableMatrix
    m = cls([[48, 49, 31],
             [ 9, 71, 94],
             [59, 28, 65]])
    assert all(type(m.inv(s)) is cls for s in 'GE ADJ LU CH LDL QR'.split())
    cls = ImmutableSparseMatrix
    m = cls([[48, 49, 31],
             [ 9, 71, 94],
             [59, 28, 65]])
    assert all(type(m.inv(s)) is cls for s in 'GE ADJ LU CH LDL QR'.split())


def test_inverse_symbolic_float_issue_26821():
    Tau, Tau_syn_in, Tau_syn_ex, C_m, Tau_syn_gap = symbols("Tau Tau_syn_in Tau_syn_ex C_m Tau_syn_gap")
    __h = symbols("__h")

    M = Matrix([
        [0,0,0,0,0,(1.0*Tau*__h-1.0*Tau_syn_in*__h)/(2.0*Tau-1.0*Tau_syn_in),-1.0*Tau*Tau_syn_in/(2.0*Tau-1.0*Tau_syn_in)],
        [0,0,0,0,0,(-1.0*Tau*__h+1.0*Tau_syn_in*__h)/(2.0*Tau*Tau_syn_in-1.0*Tau_syn_in**2),1.0],
        [0,(1.0*Tau*__h-1.0*Tau_syn_ex*__h)/(2.0*Tau-1.0*Tau_syn_ex),-1.0*Tau*Tau_syn_ex/(2.0*Tau-1.0*Tau_syn_ex),0,0,0,0],
        [0,(-1.0*Tau*__h+1.0*Tau_syn_ex*__h)/(2.0*Tau*Tau_syn_ex-1.0*Tau_syn_ex**2),1.0,0,0,0,0],
        [0,0,0,(1.0*Tau*__h-1.0*Tau_syn_gap*__h)/(2.0*Tau-1.0*Tau_syn_gap),-1.0*Tau*Tau_syn_gap/(2.0*Tau-1.0*Tau_syn_gap),0,0],
        [0,0,0,(-1.0*Tau*__h+1.0*Tau_syn_gap*__h)/(2.0*Tau*Tau_syn_gap-1.0*Tau_syn_gap**2),1.0,0,0],
        [1.0,-1.0*Tau*Tau_syn_ex*__h/(2.0*C_m*Tau-1.0*C_m*Tau_syn_ex),0,-1.0*Tau*Tau_syn_gap*__h/(2.0*C_m*Tau-1.0*C_m*Tau_syn_gap),0,-1.0*Tau*Tau_syn_in*__h/(2.0*C_m*Tau-1.0*C_m*Tau_syn_in),0]
    ])

    Mi = M.inv()

    assert (M*Mi - eye(7)).applyfunc(cancel) == zeros(7)

    # https://github.com/sympy/sympy/issues/26821
    # Previously very large floats were in the result.
    assert max(abs(f) for f in Mi.atoms(Float)) < 1e3


@slow
def test_matrix_exponential_issue_26821():
    # The symbol names matter in the original bug...
    a, b, c, d, e = symbols("Tau, Tau_syn_in, Tau_syn_ex, C_m, Tau_syn_gap")
    t = symbols("__h")
    M = Matrix([
        [      0,  1.0,       0,    0,       0,    0,    0],
        [-1/b**2, -2/b,       0,    0,       0,    0,    0],
        [      0,    0,       0,  1.0,       0,    0,    0],
        [      0,    0, -1/c**2, -2/c,       0,    0,    0],
        [      0,    0,       0,    0,       0,    1,    0],
        [      0,    0,       0,    0, -1/e**2, -2/e,    0],
        [    1/d,    0,     1/d,    0,     1/d,    0, -1/a]
    ])

    Me = (t*M).exp()
    assert (Me.diff(t) - M*Me).applyfunc(cancel) == zeros(7)
    # https://github.com/sympy/sympy/issues/26821
    # Previously very large floats were in the result.
    assert max(abs(f) for f in Me.atoms(Float)) < 1e3


def test_jacobian_hessian():
    L = Matrix(1, 2, [x**2*y, 2*y**2 + x*y])
    syms = [x, y]
    assert L.jacobian(syms) == Matrix([[2*x*y, x**2], [y, 4*y + x]])

    L = Matrix(1, 2, [x, x**2*y**3])
    assert L.jacobian(syms) == Matrix([[1, 0], [2*x*y**3, x**2*3*y**2]])

    f = x**2*y
    syms = [x, y]
    assert hessian(f, syms) == Matrix([[2*y, 2*x], [2*x, 0]])

    f = x**2*y**3
    assert hessian(f, syms) == \
        Matrix([[2*y**3, 6*x*y**2], [6*x*y**2, 6*x**2*y]])

    f = z + x*y**2
    g = x**2 + 2*y**3
    ans = Matrix([[0,   2*y],
                  [2*y, 2*x]])
    assert ans == hessian(f, Matrix([x, y]))
    assert ans == hessian(f, Matrix([x, y]).T)
    assert hessian(f, (y, x), [g]) == Matrix([
        [     0, 6*y**2, 2*x],
        [6*y**2,    2*x, 2*y],
        [   2*x,    2*y,   0]])


def test_wronskian():
    assert wronskian([cos(x), sin(x)], x) == cos(x)**2 + sin(x)**2
    assert wronskian([exp(x), exp(2*x)], x) == exp(3*x)
    assert wronskian([exp(x), x], x) == exp(x) - x*exp(x)
    assert wronskian([1, x, x**2], x) == 2
    w1 = -6*exp(x)*sin(x)*x + 6*cos(x)*exp(x)*x**2 - 6*exp(x)*cos(x)*x - \
        exp(x)*cos(x)*x**3 + exp(x)*sin(x)*x**3
    assert wronskian([exp(x), cos(x), x**3], x).expand() == w1
    assert wronskian([exp(x), cos(x), x**3], x, method='berkowitz').expand() \
        == w1
    w2 = -x**3*cos(x)**2 - x**3*sin(x)**2 - 6*x*cos(x)**2 - 6*x*sin(x)**2
    assert wronskian([sin(x), cos(x), x**3], x).expand() == w2
    assert wronskian([sin(x), cos(x), x**3], x, method='berkowitz').expand() \
        == w2
    assert wronskian([], x) == 1


def test_xreplace():
    assert Matrix([[1, x], [x, 4]]).xreplace({x: 5}) == \
        Matrix([[1, 5], [5, 4]])
    assert Matrix([[x, 2], [x + y, 4]]).xreplace({x: -1, y: -2}) == \
        Matrix([[-1, 2], [-3, 4]])
    for cls in all_classes:
        assert Matrix([[2, 0], [0, 2]]) == cls.eye(2).xreplace({1: 2})


def test_simplify():
    n = Symbol('n')
    f = Function('f')

    M = Matrix([[            1/x + 1/y,                 (x + x*y) / x  ],
                [ (f(x) + y*f(x))/f(x), 2 * (1/n - cos(n * pi)/n) / pi ]])
    M.simplify()
    assert M == Matrix([[ (x + y)/(x * y),                        1 + y ],
                        [           1 + y, 2*((1 - 1*cos(pi*n))/(pi*n)) ]])
    eq = (1 + x)**2
    M = Matrix([[eq]])
    M.simplify()
    assert M == Matrix([[eq]])
    M.simplify(ratio=oo)
    assert M == Matrix([[eq.simplify(ratio=oo)]])

    n = Symbol('n')
    f = Function('f')

    M = ImmutableMatrix([
        [            1/x + 1/y,                 (x + x*y) / x  ],
        [ (f(x) + y*f(x))/f(x), 2 * (1/n - cos(n * pi)/n) / pi ]
    ])
    assert M.simplify() == Matrix([
        [ (x + y)/(x * y),                        1 + y ],
        [           1 + y, 2*((1 - 1*cos(pi*n))/(pi*n)) ]
    ])

    eq = (1 + x)**2
    M = ImmutableMatrix([[eq]])
    assert M.simplify() == Matrix([[eq]])
    assert M.simplify(ratio=oo) == Matrix([[eq.simplify(ratio=oo)]])

    assert simplify(ImmutableMatrix([[sin(x)**2 + cos(x)**2]])) == \
                    ImmutableMatrix([[1]])

    # https://github.com/sympy/sympy/issues/19353
    m = Matrix([[30, 2], [3, 4]])
    assert (1/(m.trace())).simplify() == Rational(1, 34)

def test_transpose():
    M = Matrix([[1, 2, 3, 4, 5, 6, 7, 8, 9, 0],
                [1, 2, 3, 4, 5, 6, 7, 8, 9, 0]])
    assert M.T == Matrix( [ [1, 1],
                            [2, 2],
                            [3, 3],
                            [4, 4],
                            [5, 5],
                            [6, 6],
                            [7, 7],
                            [8, 8],
                            [9, 9],
                            [0, 0] ])
    assert M.T.T == M
    assert M.T == M.transpose()


def test_conj_dirac():
    raises(AttributeError, lambda: eye(3).D)

    M = Matrix([[1, I, I, I],
                [0, 1, I, I],
                [0, 0, 1, I],
                [0, 0, 0, 1]])

    assert M.D == Matrix([[ 1,  0,  0,  0],
                          [-I,  1,  0,  0],
                          [-I, -I, -1,  0],
                          [-I, -I,  I, -1]])


def test_trace():
    M = Matrix([[1, 0, 0],
                [0, 5, 0],
                [0, 0, 8]])
    assert M.trace() == 14


def test_shape():
    m = Matrix(1, 2, [0, 0])
    assert m.shape == (1, 2)
    M = Matrix([[x, 0, 0],
                [0, y, 0]])
    assert M.shape == (2, 3)


def test_col_row_op():
    M = Matrix([[x, 0, 0],
                [0, y, 0]])
    M.row_op(1, lambda r, j: r + j + 1)
    assert M == Matrix([[x,     0, 0],
                        [1, y + 2, 3]])

    M.col_op(0, lambda c, j: c + y**j)
    assert M == Matrix([[x + 1,     0, 0],
                        [1 + y, y + 2, 3]])

    # neither row nor slice give copies that allow the original matrix to
    # be changed
    assert M.row(0) == Matrix([[x + 1, 0, 0]])
    r1 = M.row(0)
    r1[0] = 42
    assert M[0, 0] == x + 1
    r1 = M[0, :-1]  # also testing negative slice
    r1[0] = 42
    assert M[0, 0] == x + 1
    c1 = M.col(0)
    assert c1 == Matrix([x + 1, 1 + y])
    c1[0] = 0
    assert M[0, 0] == x + 1
    c1 = M[:, 0]
    c1[0] = 42
    assert M[0, 0] == x + 1


def test_row_mult():
    M = Matrix([[1,2,3],
               [4,5,6]])
    M.row_mult(1,3)
    assert M[1,0] == 12
    assert M[0,0] == 1
    assert M[1,2] == 18


def test_row_add():
    M = Matrix([[1,2,3],
               [4,5,6],
               [1,1,1]])
    M.row_add(2,0,5)
    assert M[0,0] == 6
    assert M[1,0] == 4
    assert M[0,2] == 8


def test_zip_row_op():
    for cls in mutable_classes: # XXX: immutable matrices don't support row ops
        M = cls.eye(3)
        M.zip_row_op(1, 0, lambda v, u: v + 2*u)
        assert M == cls([[1, 0, 0],
                         [2, 1, 0],
                         [0, 0, 1]])

        M = cls.eye(3)*2
        M[0, 1] = -1
        M.zip_row_op(1, 0, lambda v, u: v + 2*u); M
        assert M == cls([[2, -1, 0],
                         [4,  0, 0],
                         [0,  0, 2]])


def test_issue_3950():
    m = Matrix([1, 2, 3])
    a = Matrix([1, 2, 3])
    b = Matrix([2, 2, 3])
    assert not (m in [])
    assert not (m in [1])
    assert m != 1
    assert m == a
    assert m != b


def test_issue_3981():
    class Index1:
        def __index__(self):
            return 1

    class Index2:
        def __index__(self):
            return 2
    index1 = Index1()
    index2 = Index2()

    m = Matrix([1, 2, 3])

    assert m[index2] == 3

    m[index2] = 5
    assert m[2] == 5

    m = Matrix([[1, 2, 3], [4, 5, 6]])
    assert m[index1, index2] == 6
    assert m[1, index2] == 6
    assert m[index1, 2] == 6

    m[index1, index2] = 4
    assert m[1, 2] == 4
    m[1, index2] = 6
    assert m[1, 2] == 6
    m[index1, 2] = 8
    assert m[1, 2] == 8


def test_is_upper():
    a = Matrix([[1, 2, 3]])
    assert a.is_upper is True
    a = Matrix([[1], [2], [3]])
    assert a.is_upper is False
    a = zeros(4, 2)
    assert a.is_upper is True


def test_is_lower():
    a = Matrix([[1, 2, 3]])
    assert a.is_lower is False
    a = Matrix([[1], [2], [3]])
    assert a.is_lower is True


def test_is_nilpotent():
    a = Matrix(4, 4, [0, 2, 1, 6, 0, 0, 1, 2, 0, 0, 0, 3, 0, 0, 0, 0])
    assert a.is_nilpotent()
    a = Matrix([[1, 0], [0, 1]])
    assert not a.is_nilpotent()
    a = Matrix([])
    assert a.is_nilpotent()


def test_zeros_ones_fill():
    n, m = 3, 5

    a = zeros(n, m)
    a.fill( 5 )

    b = 5 * ones(n, m)

    assert a == b
    assert a.rows == b.rows == 3
    assert a.cols == b.cols == 5
    assert a.shape == b.shape == (3, 5)
    assert zeros(2) == zeros(2, 2)
    assert ones(2) == ones(2, 2)
    assert zeros(2, 3) == Matrix(2, 3, [0]*6)
    assert ones(2, 3) == Matrix(2, 3, [1]*6)

    a.fill(0)
    assert a == zeros(n, m)


def test_empty_zeros():
    a = zeros(0)
    assert a == Matrix()
    a = zeros(0, 2)
    assert a.rows == 0
    assert a.cols == 2
    a = zeros(2, 0)
    assert a.rows == 2
    assert a.cols == 0


def test_issue_3749():
    a = Matrix([[x**2, x*y], [x*sin(y), x*cos(y)]])
    assert a.diff(x) == Matrix([[2*x, y], [sin(y), cos(y)]])
    assert Matrix([
        [x, -x, x**2],
        [exp(x), 1/x - exp(-x), x + 1/x]]).limit(x, oo) == \
        Matrix([[oo, -oo, oo], [oo, 0, oo]])
    assert Matrix([
        [(exp(x) - 1)/x, 2*x + y*x, x**x ],
        [1/x, abs(x), abs(sin(x + 1))]]).limit(x, 0) == \
        Matrix([[1, 0, 1], [oo, 0, sin(1)]])
    assert a.integrate(x) == Matrix([
        [Rational(1, 3)*x**3, y*x**2/2],
        [x**2*sin(y)/2, x**2*cos(y)/2]])


def test_inv_iszerofunc():
    A = eye(4)
    A.col_swap(0, 1)
    for method in "GE", "LU":
        assert A.inv(method=method, iszerofunc=lambda x: x == 0) == \
            A.inv(method="ADJ")


def test_jacobian_metrics():
    rho, phi = symbols("rho,phi")
    X = Matrix([rho*cos(phi), rho*sin(phi)])
    Y = Matrix([rho, phi])
    J = X.jacobian(Y)
    assert J == X.jacobian(Y.T)
    assert J == (X.T).jacobian(Y)
    assert J == (X.T).jacobian(Y.T)
    g = J.T*eye(J.shape[0])*J
    g = g.applyfunc(trigsimp)
    assert g == Matrix([[1, 0], [0, rho**2]])


def test_jacobian2():
    rho, phi = symbols("rho,phi")
    X = Matrix([rho*cos(phi), rho*sin(phi), rho**2])
    Y = Matrix([rho, phi])
    J = Matrix([
        [cos(phi), -rho*sin(phi)],
        [sin(phi),  rho*cos(phi)],
        [   2*rho,             0],
    ])
    assert X.jacobian(Y) == J


def test_issue_4564():
    X = Matrix([exp(x + y + z), exp(x + y + z), exp(x + y + z)])
    Y = Matrix([x, y, z])
    for i in range(1, 3):
        for j in range(1, 3):
            X_slice = X[:i, :]
            Y_slice = Y[:j, :]
            J = X_slice.jacobian(Y_slice)
            assert J.rows == i
            assert J.cols == j
            for k in range(j):
                assert J[:, k] == X_slice


def test_nonvectorJacobian():
    X = Matrix([[exp(x + y + z), exp(x + y + z)],
                [exp(x + y + z), exp(x + y + z)]])
    raises(TypeError, lambda: X.jacobian(Matrix([x, y, z])))
    X = X[0, :]
    Y = Matrix([[x, y], [x, z]])
    raises(TypeError, lambda: X.jacobian(Y))
    raises(TypeError, lambda: X.jacobian(Matrix([ [x, y], [x, z] ])))


def test_vec():
    m = Matrix([[1, 3], [2, 4]])
    m_vec = m.vec()
    assert m_vec.cols == 1
    for i in range(4):
        assert m_vec[i] == i + 1


def test_vech():
    m = Matrix([[1, 2], [2, 3]])
    m_vech = m.vech()
    assert m_vech.cols == 1
    for i in range(3):
        assert m_vech[i] == i + 1
    m_vech = m.vech(diagonal=False)
    assert m_vech[0] == 2

    m = Matrix([[1, x*(x + y)], [y*x + x**2, 1]])
    m_vech = m.vech(diagonal=False)
    assert m_vech[0] == y*x + x**2

    m = Matrix([[1, x*(x + y)], [y*x, 1]])
    m_vech = m.vech(diagonal=False, check_symmetry=False)
    assert m_vech[0] == y*x

    raises(ShapeError, lambda: Matrix([[1, 3]]).vech())
    raises(ValueError, lambda: Matrix([[1, 3], [2, 4]]).vech())
    raises(ShapeError, lambda: Matrix([[1, 3]]).vech())
    raises(ValueError, lambda: Matrix([[1, 3], [2, 4]]).vech())


def test_diag():
    # mostly tested in testcommonmatrix.py
    assert diag([1, 2, 3]) == Matrix([1, 2, 3])
    m = [1, 2, [3]]
    raises(ValueError, lambda: diag(m))
    assert diag(m, strict=False) == Matrix([1, 2, 3])


def test_inv_block():
    a = Matrix([[1, 2], [2, 3]])
    b = Matrix([[3, x], [y, 3]])
    c = Matrix([[3, x, 3], [y, 3, z], [x, y, z]])
    A = diag(a, b, b)
    assert A.inv(try_block_diag=True) == diag(a.inv(), b.inv(), b.inv())
    A = diag(a, b, c)
    assert A.inv(try_block_diag=True) == diag(a.inv(), b.inv(), c.inv())
    A = diag(a, c, b)
    assert A.inv(try_block_diag=True) == diag(a.inv(), c.inv(), b.inv())
    A = diag(a, a, b, a, c, a)
    assert A.inv(try_block_diag=True) == diag(
        a.inv(), a.inv(), b.inv(), a.inv(), c.inv(), a.inv())
    assert A.inv(try_block_diag=True, method="ADJ") == diag(
        a.inv(method="ADJ"), a.inv(method="ADJ"), b.inv(method="ADJ"),
        a.inv(method="ADJ"), c.inv(method="ADJ"), a.inv(method="ADJ"))


def test_creation_args():
    """
    Check that matrix dimensions can be specified using any reasonable type
    (see issue 4614).
    """
    raises(ValueError, lambda: zeros(3, -1))
    raises(TypeError, lambda: zeros(1, 2, 3, 4))
    assert zeros(int(3)) == zeros(3)
    assert zeros(Integer(3)) == zeros(3)
    raises(ValueError, lambda: zeros(3.))
    assert eye(int(3)) == eye(3)
    assert eye(Integer(3)) == eye(3)
    raises(ValueError, lambda: eye(3.))
    assert ones(int(3), Integer(4)) == ones(3, 4)
    raises(TypeError, lambda: Matrix(5))
    raises(TypeError, lambda: Matrix(1, 2))
    raises(ValueError, lambda: Matrix([1, [2]]))


def test_diagonal_symmetrical():
    m = Matrix(2, 2, [0, 1, 1, 0])
    assert not m.is_diagonal()
    assert m.is_symmetric()
    assert m.is_symmetric(simplify=False)

    m = Matrix(2, 2, [1, 0, 0, 1])
    assert m.is_diagonal()

    m = diag(1, 2, 3)
    assert m.is_diagonal()
    assert m.is_symmetric()

    m = Matrix(3, 3, [1, 0, 0, 0, 2, 0, 0, 0, 3])
    assert m == diag(1, 2, 3)

    m = Matrix(2, 3, zeros(2, 3))
    assert not m.is_symmetric()
    assert m.is_diagonal()

    m = Matrix(((5, 0), (0, 6), (0, 0)))
    assert m.is_diagonal()

    m = Matrix(((5, 0, 0), (0, 6, 0)))
    assert m.is_diagonal()

    m = Matrix(3, 3, [1, x**2 + 2*x + 1, y, (x + 1)**2, 2, 0, y, 0, 3])
    assert m.is_symmetric()
    assert not m.is_symmetric(simplify=False)
    assert m.expand().is_symmetric(simplify=False)


def test_diagonalization():
    m = Matrix([[1, 2+I], [2-I, 3]])
    assert m.is_diagonalizable()

    m = Matrix(3, 2, [-3, 1, -3, 20, 3, 10])
    assert not m.is_diagonalizable()
    assert not m.is_symmetric()
    raises(NonSquareMatrixError, lambda: m.diagonalize())

    # diagonalizable
    m = diag(1, 2, 3)
    (P, D) = m.diagonalize()
    assert P == eye(3)
    assert D == m

    m = Matrix(2, 2, [0, 1, 1, 0])
    assert m.is_symmetric()
    assert m.is_diagonalizable()
    (P, D) = m.diagonalize()
    assert P.inv() * m * P == D

    m = Matrix(2, 2, [1, 0, 0, 3])
    assert m.is_symmetric()
    assert m.is_diagonalizable()
    (P, D) = m.diagonalize()
    assert P.inv() * m * P == D
    assert P == eye(2)
    assert D == m

    m = Matrix(2, 2, [1, 1, 0, 0])
    assert m.is_diagonalizable()
    (P, D) = m.diagonalize()
    assert P.inv() * m * P == D

    m = Matrix(3, 3, [1, 2, 0, 0, 3, 0, 2, -4, 2])
    assert m.is_diagonalizable()
    (P, D) = m.diagonalize()
    assert P.inv() * m * P == D
    for i in P:
        assert i.as_numer_denom()[1] == 1

    m = Matrix(2, 2, [1, 0, 0, 0])
    assert m.is_diagonal()
    assert m.is_diagonalizable()
    (P, D) = m.diagonalize()
    assert P.inv() * m * P == D
    assert P == Matrix([[0, 1], [1, 0]])

    # diagonalizable, complex only
    m = Matrix(2, 2, [0, 1, -1, 0])
    assert not m.is_diagonalizable(True)
    raises(MatrixError, lambda: m.diagonalize(True))
    assert m.is_diagonalizable()
    (P, D) = m.diagonalize()
    assert P.inv() * m * P == D

    # not diagonalizable
    m = Matrix(2, 2, [0, 1, 0, 0])
    assert not m.is_diagonalizable()
    raises(MatrixError, lambda: m.diagonalize())

    m = Matrix(3, 3, [-3, 1, -3, 20, 3, 10, 2, -2, 4])
    assert not m.is_diagonalizable()
    raises(MatrixError, lambda: m.diagonalize())

    # symbolic
    a, b, c, d = symbols('a b c d')
    m = Matrix(2, 2, [a, c, c, b])
    assert m.is_symmetric()
    assert m.is_diagonalizable()


def test_issue_15887():
    # Mutable matrix should not use cache
    a = MutableDenseMatrix([[0, 1], [1, 0]])
    assert a.is_diagonalizable() is True
    a[1, 0] = 0
    assert a.is_diagonalizable() is False

    a = MutableDenseMatrix([[0, 1], [1, 0]])
    a.diagonalize()
    a[1, 0] = 0
    raises(MatrixError, lambda: a.diagonalize())


def test_jordan_form():

    m = Matrix(3, 2, [-3, 1, -3, 20, 3, 10])
    raises(NonSquareMatrixError, lambda: m.jordan_form())

    # diagonalizable
    m = Matrix(3, 3, [7, -12, 6, 10, -19, 10, 12, -24, 13])
    Jmust = Matrix(3, 3, [-1, 0, 0, 0, 1, 0, 0, 0, 1])
    P, J = m.jordan_form()
    assert Jmust == J
    assert Jmust == m.diagonalize()[1]

    # m = Matrix(3, 3, [0, 6, 3, 1, 3, 1, -2, 2, 1])
    # m.jordan_form()  # very long
    # m.jordan_form()  #

    # diagonalizable, complex only

    # Jordan cells
    # complexity: one of eigenvalues is zero
    m = Matrix(3, 3, [0, 1, 0, -4, 4, 0, -2, 1, 2])
    # The blocks are ordered according to the value of their eigenvalues,
    # in order to make the matrix compatible with .diagonalize()
    Jmust = Matrix(3, 3, [2, 1, 0, 0, 2, 0, 0, 0, 2])
    P, J = m.jordan_form()
    assert Jmust == J

    # complexity: all of eigenvalues are equal
    m = Matrix(3, 3, [2, 6, -15, 1, 1, -5, 1, 2, -6])
    # Jmust = Matrix(3, 3, [-1, 0, 0, 0, -1, 1, 0, 0, -1])
    # same here see 1456ff
    Jmust = Matrix(3, 3, [-1, 1, 0, 0, -1, 0, 0, 0, -1])
    P, J = m.jordan_form()
    assert Jmust == J

    # complexity: two of eigenvalues are zero
    m = Matrix(3, 3, [4, -5, 2, 5, -7, 3, 6, -9, 4])
    Jmust = Matrix(3, 3, [0, 1, 0, 0, 0, 0, 0, 0, 1])
    P, J = m.jordan_form()
    assert Jmust == J

    m = Matrix(4, 4, [6, 5, -2, -3, -3, -1, 3, 3, 2, 1, -2, -3, -1, 1, 5, 5])
    Jmust = Matrix(4, 4, [2, 1, 0, 0,
                          0, 2, 0, 0,
              0, 0, 2, 1,
              0, 0, 0, 2]
              )
    P, J = m.jordan_form()
    assert Jmust == J

    m = Matrix(4, 4, [6, 2, -8, -6, -3, 2, 9, 6, 2, -2, -8, -6, -1, 0, 3, 4])
    # Jmust = Matrix(4, 4, [2, 0, 0, 0, 0, 2, 1, 0, 0, 0, 2, 0, 0, 0, 0, -2])
    # same here see 1456ff
    Jmust = Matrix(4, 4, [-2, 0, 0, 0,
                           0, 2, 1, 0,
                           0, 0, 2, 0,
                           0, 0, 0, 2])
    P, J = m.jordan_form()
    assert Jmust == J

    m = Matrix(4, 4, [5, 4, 2, 1, 0, 1, -1, -1, -1, -1, 3, 0, 1, 1, -1, 2])
    assert not m.is_diagonalizable()
    Jmust = Matrix(4, 4, [1, 0, 0, 0, 0, 2, 0, 0, 0, 0, 4, 1, 0, 0, 0, 4])
    P, J = m.jordan_form()
    assert Jmust == J

    # checking for maximum precision to remain unchanged
    m = Matrix([[Float('1.0', precision=110), Float('2.0', precision=110)],
                [Float('3.14159265358979323846264338327', precision=110), Float('4.0', precision=110)]])
    P, J = m.jordan_form()
    for term in J.values():
        if isinstance(term, Float):
            assert term._prec == 110


def test_jordan_form_complex_issue_9274():
    A = Matrix([[ 2,  4,  1,  0],
                [-4,  2,  0,  1],
                [ 0,  0,  2,  4],
                [ 0,  0, -4,  2]])
    p = 2 - 4*I
    q = 2 + 4*I
    Jmust1 = Matrix([[p, 1, 0, 0],
                     [0, p, 0, 0],
                     [0, 0, q, 1],
                     [0, 0, 0, q]])
    Jmust2 = Matrix([[q, 1, 0, 0],
                     [0, q, 0, 0],
                     [0, 0, p, 1],
                     [0, 0, 0, p]])
    P, J = A.jordan_form()
    assert J == Jmust1 or J == Jmust2
    assert simplify(P*J*P.inv()) == A


def test_issue_10220():
    # two non-orthogonal Jordan blocks with eigenvalue 1
    M = Matrix([[1, 0, 0, 1],
                [0, 1, 1, 0],
                [0, 0, 1, 1],
                [0, 0, 0, 1]])
    P, J = M.jordan_form()
    assert P == Matrix([[0, 1, 0, 1],
                        [1, 0, 0, 0],
                        [0, 1, 0, 0],
                        [0, 0, 1, 0]])
    assert J == Matrix([
                        [1, 1, 0, 0],
                        [0, 1, 1, 0],
                        [0, 0, 1, 0],
                        [0, 0, 0, 1]])


def test_jordan_form_issue_15858():
    A = Matrix([
        [1, 1, 1, 0],
        [-2, -1, 0, -1],
        [0, 0, -1, -1],
        [0, 0, 2, 1]])
    (P, J) = A.jordan_form()
    assert P.expand() == Matrix([
        [    -I,          -I/2,      I,           I/2],
        [-1 + I,             0, -1 - I,             0],
        [     0, -S(1)/2 - I/2,      0, -S(1)/2 + I/2],
        [     0,             1,      0,             1]])
    assert J == Matrix([
        [-I, 1, 0, 0],
        [0, -I, 0, 0],
        [0, 0, I, 1],
        [0, 0, 0, I]])


def test_Matrix_berkowitz_charpoly():
    UA, K_i, K_w = symbols('UA K_i K_w')

    A = Matrix([[-K_i - UA + K_i**2/(K_i + K_w),       K_i*K_w/(K_i + K_w)],
                [           K_i*K_w/(K_i + K_w), -K_w + K_w**2/(K_i + K_w)]])

    charpoly = A.charpoly(x)

    assert charpoly == \
        Poly(x**2 + (K_i*UA + K_w*UA + 2*K_i*K_w)/(K_i + K_w)*x +
        K_i*K_w*UA/(K_i + K_w), x, domain='ZZ(K_i,K_w,UA)')

    assert type(charpoly) is PurePoly

    A = Matrix([[1, 3], [2, 0]])
    assert A.charpoly() == A.charpoly(x) == PurePoly(x**2 - x - 6)

    A = Matrix([[1, 2], [x, 0]])
    p = A.charpoly(x)
    assert p.gen != x
    assert p.as_expr().subs(p.gen, x) == x**2 - 3*x


def test_exp_jordan_block():
    l = Symbol('lamda')

    m = Matrix.jordan_block(1, l)
    assert m._eval_matrix_exp_jblock() == Matrix([[exp(l)]])

    m = Matrix.jordan_block(3, l)
    assert m._eval_matrix_exp_jblock() == \
        Matrix([
            [exp(l), exp(l), exp(l)/2],
            [0, exp(l), exp(l)],
            [0, 0, exp(l)]])


def test_exp():
    m = Matrix([[3, 4], [0, -2]])
    m_exp = Matrix([[exp(3), -4*exp(-2)/5 + 4*exp(3)/5], [0, exp(-2)]])
    assert m.exp() == m_exp
    assert exp(m) == m_exp

    m = Matrix([[1, 0], [0, 1]])
    assert m.exp() == Matrix([[E, 0], [0, E]])
    assert exp(m) == Matrix([[E, 0], [0, E]])

    m = Matrix([[1, -1], [1, 1]])
    assert m.exp() == Matrix([[E*cos(1), -E*sin(1)], [E*sin(1), E*cos(1)]])


def test_log():
    l = Symbol('lamda')

    m = Matrix.jordan_block(1, l)
    assert m._eval_matrix_log_jblock() == Matrix([[log(l)]])

    m = Matrix.jordan_block(4, l)
    assert m._eval_matrix_log_jblock() == \
        Matrix(
            [
                [log(l), 1/l, -1/(2*l**2), 1/(3*l**3)],
                [0, log(l), 1/l, -1/(2*l**2)],
                [0, 0, log(l), 1/l],
                [0, 0, 0, log(l)]
            ]
        )

    m = Matrix(
        [[0, 0, 1],
        [0, 0, 0],
        [-1, 0, 0]]
    )
    raises(MatrixError, lambda: m.log())


def test_find_reasonable_pivot_naive_finds_guaranteed_nonzero1():
    # Test if matrices._find_reasonable_pivot_naive()
    # finds a guaranteed non-zero pivot when the
    # some of the candidate pivots are symbolic expressions.
    # Keyword argument: simpfunc=None indicates that no simplifications
    # should be performed during the search.
    x = Symbol('x')
    column = Matrix(3, 1, [x, cos(x)**2 + sin(x)**2, S.Half])
    pivot_offset, pivot_val, pivot_assumed_nonzero, simplified =\
        _find_reasonable_pivot_naive(column)
    assert pivot_val == S.Half


def test_find_reasonable_pivot_naive_finds_guaranteed_nonzero2():
    # Test if matrices._find_reasonable_pivot_naive()
    # finds a guaranteed non-zero pivot when the
    # some of the candidate pivots are symbolic expressions.
    # Keyword argument: simpfunc=_simplify indicates that the search
    # should attempt to simplify candidate pivots.
    x = Symbol('x')
    column = Matrix(3, 1,
                    [x,
                     cos(x)**2+sin(x)**2+x**2,
                     cos(x)**2+sin(x)**2])
    pivot_offset, pivot_val, pivot_assumed_nonzero, simplified =\
        _find_reasonable_pivot_naive(column, simpfunc=_simplify)
    assert pivot_val == 1


def test_find_reasonable_pivot_naive_simplifies():
    # Test if matrices._find_reasonable_pivot_naive()
    # simplifies candidate pivots, and reports
    # their offsets correctly.
    x = Symbol('x')
    column = Matrix(3, 1,
                    [x,
                     cos(x)**2+sin(x)**2+x,
                     cos(x)**2+sin(x)**2])
    pivot_offset, pivot_val, pivot_assumed_nonzero, simplified =\
        _find_reasonable_pivot_naive(column, simpfunc=_simplify)

    assert len(simplified) == 2
    assert simplified[0][0] == 1
    assert simplified[0][1] == 1+x
    assert simplified[1][0] == 2
    assert simplified[1][1] == 1


def test_errors():
    raises(ValueError, lambda: Matrix([[1, 2], [1]]))
    raises(IndexError, lambda: Matrix([[1, 2]])[1.2, 5])
    raises(IndexError, lambda: Matrix([[1, 2]])[1, 5.2])
    raises(ValueError, lambda: randMatrix(3, c=4, symmetric=True))
    raises(ValueError, lambda: Matrix([1, 2]).reshape(4, 6))
    raises(ShapeError,
        lambda: Matrix([[1, 2], [3, 4]]).copyin_matrix([1, 0], Matrix([1, 2])))
    raises(TypeError, lambda: Matrix([[1, 2], [3, 4]]).copyin_list([0,
           1], set()))
    raises(NonSquareMatrixError, lambda: Matrix([[1, 2, 3], [2, 3, 0]]).inv())
    raises(ShapeError,
        lambda: Matrix(1, 2, [1, 2]).row_join(Matrix([[1, 2], [3, 4]])))
    raises(
        ShapeError, lambda: Matrix([1, 2]).col_join(Matrix([[1, 2], [3, 4]])))
    raises(ShapeError, lambda: Matrix([1]).row_insert(1, Matrix([[1,
           2], [3, 4]])))
    raises(ShapeError, lambda: Matrix([1]).col_insert(1, Matrix([[1,
           2], [3, 4]])))
    raises(NonSquareMatrixError, lambda: Matrix([1, 2]).trace())
    raises(TypeError, lambda: Matrix([1]).applyfunc(1))
    raises(ValueError, lambda: Matrix([[1, 2], [3, 4]]).minor(4, 5))
    raises(ValueError, lambda: Matrix([[1, 2], [3, 4]]).minor_submatrix(4, 5))
    raises(TypeError, lambda: Matrix([1, 2, 3]).cross(1))
    raises(TypeError, lambda: Matrix([1, 2, 3]).dot(1))
    raises(ShapeError, lambda: Matrix([1, 2, 3]).dot(Matrix([1, 2])))
    raises(ShapeError, lambda: Matrix([1, 2]).dot([]))
    raises(TypeError, lambda: Matrix([1, 2]).dot('a'))
    raises(ShapeError, lambda: Matrix([1, 2]).dot([1, 2, 3]))
    raises(NonSquareMatrixError, lambda: Matrix([1, 2, 3]).exp())
    raises(ShapeError, lambda: Matrix([[1, 2], [3, 4]]).normalized())
    raises(ValueError, lambda: Matrix([1, 2]).inv(method='not a method'))
    raises(NonSquareMatrixError, lambda: Matrix([1, 2]).inverse_GE())
    raises(ValueError, lambda: Matrix([[1, 2], [1, 2]]).inverse_GE())
    raises(NonSquareMatrixError, lambda: Matrix([1, 2]).inverse_ADJ())
    raises(ValueError, lambda: Matrix([[1, 2], [1, 2]]).inverse_ADJ())
    raises(NonSquareMatrixError, lambda: Matrix([1, 2]).inverse_LU())
    raises(NonSquareMatrixError, lambda: Matrix([1, 2]).is_nilpotent())
    raises(NonSquareMatrixError, lambda: Matrix([1, 2]).det())
    raises(ValueError,
        lambda: Matrix([[1, 2], [3, 4]]).det(method='Not a real method'))
    raises(ValueError,
        lambda: Matrix([[1, 2, 3, 4], [5, 6, 7, 8],
        [9, 10, 11, 12], [13, 14, 15, 16]]).det(iszerofunc="Not function"))
    raises(ValueError,
        lambda: Matrix([[1, 2, 3, 4], [5, 6, 7, 8],
        [9, 10, 11, 12], [13, 14, 15, 16]]).det(iszerofunc=False))
    raises(ValueError,
        lambda: hessian(Matrix([[1, 2], [3, 4]]), Matrix([[1, 2], [2, 1]])))
    raises(ValueError, lambda: hessian(Matrix([[1, 2], [3, 4]]), []))
    raises(ValueError, lambda: hessian(Symbol('x')**2, 'a'))
    raises(IndexError, lambda: eye(3)[5, 2])
    raises(IndexError, lambda: eye(3)[2, 5])
    M = Matrix(((1, 2, 3, 4), (5, 6, 7, 8), (9, 10, 11, 12), (13, 14, 15, 16)))
    raises(ValueError, lambda: M.det('method=LU_decomposition()'))
    V = Matrix([[10, 10, 10]])
    M = Matrix([[1, 2, 3], [2, 3, 4], [3, 4, 5]])
    raises(ValueError, lambda: M.row_insert(4.7, V))
    M = Matrix([[1, 2, 3], [2, 3, 4], [3, 4, 5]])
    raises(ValueError, lambda: M.col_insert(-4.2, V))


def test_len():
    assert len(Matrix()) == 0
    assert len(Matrix([[1, 2]])) == len(Matrix([[1], [2]])) == 2
    assert len(Matrix(0, 2, lambda i, j: 0)) == \
        len(Matrix(2, 0, lambda i, j: 0)) == 0
    assert len(Matrix([[0, 1, 2], [3, 4, 5]])) == 6
    assert Matrix([1]) == Matrix([[1]])
    assert not Matrix()
    assert Matrix() == Matrix([])


def test_integrate():
    A = Matrix(((1, 4, x), (y, 2, 4), (10, 5, x**2)))
    assert A.integrate(x) == \
        Matrix(((x, 4*x, x**2/2), (x*y, 2*x, 4*x), (10*x, 5*x, x**3/3)))
    assert A.integrate(y) == \
        Matrix(((y, 4*y, x*y), (y**2/2, 2*y, 4*y), (10*y, 5*y, y*x**2)))
    m = Matrix(2, 1, [x, y])
    assert m.integrate(x) == Matrix(2, 1, [x**2/2, y*x])


def test_diff():
    A = MutableDenseMatrix(((1, 4, x), (y, 2, 4), (10, 5, x**2 + 1)))
    assert isinstance(A.diff(x), type(A))
    assert A.diff(x) == MutableDenseMatrix(((0, 0, 1), (0, 0, 0), (0, 0, 2*x)))
    assert A.diff(y) == MutableDenseMatrix(((0, 0, 0), (1, 0, 0), (0, 0, 0)))

    assert diff(A, x) == MutableDenseMatrix(((0, 0, 1), (0, 0, 0), (0, 0, 2*x)))
    assert diff(A, y) == MutableDenseMatrix(((0, 0, 0), (1, 0, 0), (0, 0, 0)))

    A_imm = A.as_immutable()
    assert isinstance(A_imm.diff(x), type(A_imm))
    assert A_imm.diff(x) == ImmutableDenseMatrix(((0, 0, 1), (0, 0, 0), (0, 0, 2*x)))
    assert A_imm.diff(y) == ImmutableDenseMatrix(((0, 0, 0), (1, 0, 0), (0, 0, 0)))

    assert diff(A_imm, x) == ImmutableDenseMatrix(((0, 0, 1), (0, 0, 0), (0, 0, 2*x)))
    assert diff(A_imm, y) == ImmutableDenseMatrix(((0, 0, 0), (1, 0, 0), (0, 0, 0)))

    assert A.diff(x, evaluate=False) == ArrayDerivative(A, x, evaluate=False)
    assert diff(A, x, evaluate=False) == ArrayDerivative(A, x, evaluate=False)


def test_diff_by_matrix():

    # Derive matrix by matrix:

    A = MutableDenseMatrix([[x, y], [z, t]])
    assert A.diff(A) == Array([[[[1, 0], [0, 0]], [[0, 1], [0, 0]]], [[[0, 0], [1, 0]], [[0, 0], [0, 1]]]])
    assert diff(A, A) == Array([[[[1, 0], [0, 0]], [[0, 1], [0, 0]]], [[[0, 0], [1, 0]], [[0, 0], [0, 1]]]])

    A_imm = A.as_immutable()
    assert A_imm.diff(A_imm) == Array([[[[1, 0], [0, 0]], [[0, 1], [0, 0]]], [[[0, 0], [1, 0]], [[0, 0], [0, 1]]]])
    assert diff(A_imm, A_imm) == Array([[[[1, 0], [0, 0]], [[0, 1], [0, 0]]], [[[0, 0], [1, 0]], [[0, 0], [0, 1]]]])

    # Derive a constant matrix:
    assert A.diff(a) == MutableDenseMatrix([[0, 0], [0, 0]])

    B = ImmutableDenseMatrix([a, b])
    assert A.diff(B) == Array.zeros(2, 1, 2, 2)
    assert A.diff(A) == Array([[[[1, 0], [0, 0]], [[0, 1], [0, 0]]], [[[0, 0], [1, 0]], [[0, 0], [0, 1]]]])

    # Test diff with tuples:

    dB = B.diff([[a, b]])
    assert dB.shape == (2, 2, 1)
    assert dB == Array([[[1], [0]], [[0], [1]]])

    f = Function("f")
    fxyz = f(x, y, z)
    assert fxyz.diff([[x, y, z]]) == Array([fxyz.diff(x), fxyz.diff(y), fxyz.diff(z)])
    assert fxyz.diff(([x, y, z], 2)) == Array([
        [fxyz.diff(x, 2), fxyz.diff(x, y), fxyz.diff(x, z)],
        [fxyz.diff(x, y), fxyz.diff(y, 2), fxyz.diff(y, z)],
        [fxyz.diff(x, z), fxyz.diff(z, y), fxyz.diff(z, 2)],
    ])

    expr = sin(x)*exp(y)
    assert expr.diff([[x, y]]) == Array([cos(x)*exp(y), sin(x)*exp(y)])
    assert expr.diff(y, ((x, y),)) == Array([cos(x)*exp(y), sin(x)*exp(y)])
    assert expr.diff(x, ((x, y),)) == Array([-sin(x)*exp(y), cos(x)*exp(y)])
    assert expr.diff(((y, x),), [[x, y]]) == Array([[cos(x)*exp(y), -sin(x)*exp(y)], [sin(x)*exp(y), cos(x)*exp(y)]])

    # Test different notations:

    assert fxyz.diff(x).diff(y).diff(x) == fxyz.diff(((x, y, z),), 3)[0, 1, 0]
    assert fxyz.diff(z).diff(y).diff(x) == fxyz.diff(((x, y, z),), 3)[2, 1, 0]
    assert fxyz.diff([[x, y, z]], ((z, y, x),)) == Array([[fxyz.diff(i).diff(j) for i in (x, y, z)] for j in (z, y, x)])

    # Test scalar derived by matrix remains matrix:
    res = x.diff(Matrix([[x, y]]))
    assert isinstance(res, ImmutableDenseMatrix)
    assert res == Matrix([[1, 0]])
    res = (x**3).diff(Matrix([[x, y]]))
    assert isinstance(res, ImmutableDenseMatrix)
    assert res == Matrix([[3*x**2, 0]])


def test_getattr():
    A = Matrix(((1, 4, x), (y, 2, 4), (10, 5, x**2 + 1)))
    raises(AttributeError, lambda: A.nonexistantattribute)
    assert getattr(A, 'diff')(x) == Matrix(((0, 0, 1), (0, 0, 0), (0, 0, 2*x)))


def test_hessenberg():
    A = Matrix([[3, 4, 1], [2, 4, 5], [0, 1, 2]])
    assert A.is_upper_hessenberg
    A = A.T
    assert A.is_lower_hessenberg
    A[0, -1] = 1
    assert A.is_lower_hessenberg is False

    A = Matrix([[3, 4, 1], [2, 4, 5], [3, 1, 2]])
    assert not A.is_upper_hessenberg

    A = zeros(5, 2)
    assert A.is_upper_hessenberg


def test_cholesky():
    raises(NonSquareMatrixError, lambda: Matrix((1, 2)).cholesky())
    raises(ValueError, lambda: Matrix(((1, 2), (3, 4))).cholesky())
    raises(ValueError, lambda: Matrix(((5 + I, 0), (0, 1))).cholesky())
    raises(ValueError, lambda: Matrix(((1, 5), (5, 1))).cholesky())
    raises(ValueError, lambda: Matrix(((1, 2), (3, 4))).cholesky(hermitian=False))
    assert Matrix(((5 + I, 0), (0, 1))).cholesky(hermitian=False) == Matrix([
        [sqrt(5 + I), 0], [0, 1]])
    A = Matrix(((1, 5), (5, 1)))
    L = A.cholesky(hermitian=False)
    assert L == Matrix([[1, 0], [5, 2*sqrt(6)*I]])
    assert L*L.T == A
    A = Matrix(((25, 15, -5), (15, 18, 0), (-5, 0, 11)))
    L = A.cholesky()
    assert L * L.T == A
    assert L.is_lower
    assert L == Matrix([[5, 0, 0], [3, 3, 0], [-1, 1, 3]])
    A = Matrix(((4, -2*I, 2 + 2*I), (2*I, 2, -1 + I), (2 - 2*I, -1 - I, 11)))
    assert A.cholesky().expand() == Matrix(((2, 0, 0), (I, 1, 0), (1 - I, 0, 3)))

    raises(NonSquareMatrixError, lambda: SparseMatrix((1, 2)).cholesky())
    raises(ValueError, lambda: SparseMatrix(((1, 2), (3, 4))).cholesky())
    raises(ValueError, lambda: SparseMatrix(((5 + I, 0), (0, 1))).cholesky())
    raises(ValueError, lambda: SparseMatrix(((1, 5), (5, 1))).cholesky())
    raises(ValueError, lambda: SparseMatrix(((1, 2), (3, 4))).cholesky(hermitian=False))
    assert SparseMatrix(((5 + I, 0), (0, 1))).cholesky(hermitian=False) == Matrix([
        [sqrt(5 + I), 0], [0, 1]])
    A = SparseMatrix(((1, 5), (5, 1)))
    L = A.cholesky(hermitian=False)
    assert L == Matrix([[1, 0], [5, 2*sqrt(6)*I]])
    assert L*L.T == A
    A = SparseMatrix(((25, 15, -5), (15, 18, 0), (-5, 0, 11)))
    L = A.cholesky()
    assert L * L.T == A
    assert L.is_lower
    assert L == Matrix([[5, 0, 0], [3, 3, 0], [-1, 1, 3]])
    A = SparseMatrix(((4, -2*I, 2 + 2*I), (2*I, 2, -1 + I), (2 - 2*I, -1 - I, 11)))
    assert A.cholesky() == Matrix(((2, 0, 0), (I, 1, 0), (1 - I, 0, 3)))


def test_matrix_norm():
    # Vector Tests
    # Test columns and symbols
    x = Symbol('x', real=True)
    v = Matrix([cos(x), sin(x)])
    assert trigsimp(v.norm(2)) == 1
    assert v.norm(10) == Pow(cos(x)**10 + sin(x)**10, Rational(1, 10))

    # Test Rows
    A = Matrix([[5, Rational(3, 2)]])
    assert A.norm() == Pow(25 + Rational(9, 4), S.Half)
    assert A.norm(oo) == max(A)
    assert A.norm(-oo) == min(A)

    # Matrix Tests
    # Intuitive test
    A = Matrix([[1, 1], [1, 1]])
    assert A.norm(2) == 2
    assert A.norm(-2) == 0
    assert A.norm('frobenius') == 2
    assert eye(10).norm(2) == eye(10).norm(-2) == 1
    assert A.norm(oo) == 2

    # Test with Symbols and more complex entries
    A = Matrix([[3, y, y], [x, S.Half, -pi]])
    assert (A.norm('fro')
           == sqrt(Rational(37, 4) + 2*abs(y)**2 + pi**2 + x**2))

    # Check non-square
    A = Matrix([[1, 2, -3], [4, 5, Rational(13, 2)]])
    assert A.norm(2) == sqrt(Rational(389, 8) + sqrt(78665)/8)
    assert A.norm(-2) is S.Zero
    assert A.norm('frobenius') == sqrt(389)/2

    # Test properties of matrix norms
    # https://en.wikipedia.org/wiki/Matrix_norm#Definition
    # Two matrices
    A = Matrix([[1, 2], [3, 4]])
    B = Matrix([[5, 5], [-2, 2]])
    C = Matrix([[0, -I], [I, 0]])
    D = Matrix([[1, 0], [0, -1]])
    L = [A, B, C, D]
    alpha = Symbol('alpha', real=True)

    for order in ['fro', 2, -2]:
        # Zero Check
        assert zeros(3).norm(order) is S.Zero
        # Check Triangle Inequality for all Pairs of Matrices
        for X in L:
            for Y in L:
                dif = (X.norm(order) + Y.norm(order) -
                    (X + Y).norm(order))
                assert (dif >= 0)
        # Scalar multiplication linearity
        for M in [A, B, C, D]:
            dif = simplify((alpha*M).norm(order) -
                    abs(alpha) * M.norm(order))
            assert dif == 0

    # Test Properties of Vector Norms
    # https://en.wikipedia.org/wiki/Vector_norm
    # Two column vectors
    a = Matrix([1, 1 - 1*I, -3])
    b = Matrix([S.Half, 1*I, 1])
    c = Matrix([-1, -1, -1])
    d = Matrix([3, 2, I])
    e = Matrix([Integer(1e2), Rational(1, 1e2), 1])
    L = [a, b, c, d, e]
    alpha = Symbol('alpha', real=True)

    for order in [1, 2, -1, -2, S.Infinity, S.NegativeInfinity, pi]:
        # Zero Check
        if order > 0:
            assert Matrix([0, 0, 0]).norm(order) is S.Zero
        # Triangle inequality on all pairs
        if order >= 1:  # Triangle InEq holds only for these norms
            for X in L:
                for Y in L:
                    dif = (X.norm(order) + Y.norm(order) -
                        (X + Y).norm(order))
                    assert simplify(dif >= 0) is S.true
        # Linear to scalar multiplication
        if order in [1, 2, -1, -2, S.Infinity, S.NegativeInfinity]:
            for X in L:
                dif = simplify((alpha*X).norm(order) -
                    (abs(alpha) * X.norm(order)))
                assert dif == 0

    # ord=1
    M = Matrix(3, 3, [1, 3, 0, -2, -1, 0, 3, 9, 6])
    assert M.norm(1) == 13


def test_condition_number():
    x = Symbol('x', real=True)
    A = eye(3)
    A[0, 0] = 10
    A[2, 2] = Rational(1, 10)
    assert A.condition_number() == 100

    A[1, 1] = x
    assert A.condition_number() == Max(10, Abs(x)) / Min(Rational(1, 10), Abs(x))

    M = Matrix([[cos(x), sin(x)], [-sin(x), cos(x)]])
    Mc = M.condition_number()
    assert all(Float(1.).epsilon_eq(Mc.subs(x, val).evalf()) for val in
        [Rational(1, 5), S.Half, Rational(1, 10), pi/2, pi, pi*Rational(7, 4) ])

    #issue 10782
    assert Matrix([]).condition_number() == 0


def test_equality():
    A = Matrix(((1, 2, 3), (4, 5, 6), (7, 8, 9)))
    B = Matrix(((9, 8, 7), (6, 5, 4), (3, 2, 1)))
    assert A == A[:, :]
    assert not A != A[:, :]
    assert not A == B
    assert A != B
    assert A != 10
    assert not A == 10

    # A SparseMatrix can be equal to a Matrix
    C = SparseMatrix(((1, 0, 0), (0, 1, 0), (0, 0, 1)))
    D = Matrix(((1, 0, 0), (0, 1, 0), (0, 0, 1)))
    assert C == D
    assert not C != D


def test_normalized():
    assert Matrix([3, 4]).normalized() == \
        Matrix([Rational(3, 5), Rational(4, 5)])

    # Zero vector trivial cases
    assert Matrix([0, 0, 0]).normalized() == Matrix([0, 0, 0])

    # Machine precision error truncation trivial cases
    m = Matrix([0,0,1.e-100])
    assert m.normalized(
    iszerofunc=lambda x: x.evalf(n=10, chop=True).is_zero
    ) == Matrix([0, 0, 0])


def test_print_nonzero():
    assert capture(lambda: eye(3).print_nonzero()) == \
        '[X  ]\n[ X ]\n[  X]\n'
    assert capture(lambda: eye(3).print_nonzero('.')) == \
        '[.  ]\n[ . ]\n[  .]\n'


def test_zeros_eye():
    assert Matrix.eye(3) == eye(3)
    assert Matrix.zeros(3) == zeros(3)
    assert ones(3, 4) == Matrix(3, 4, [1]*12)

    i = Matrix([[1, 0], [0, 1]])
    z = Matrix([[0, 0], [0, 0]])
    for cls in all_classes:
        m = cls.eye(2)
        assert i == m  # but m == i will fail if m is immutable
        assert i == eye(2, cls=cls)
        assert type(m) == cls
        m = cls.zeros(2)
        assert z == m
        assert z == zeros(2, cls=cls)
        assert type(m) == cls


def test_is_zero():
    assert Matrix().is_zero_matrix
    assert Matrix([[0, 0], [0, 0]]).is_zero_matrix
    assert zeros(3, 4).is_zero_matrix
    assert not eye(3).is_zero_matrix
    assert Matrix([[x, 0], [0, 0]]).is_zero_matrix == None
    assert SparseMatrix([[x, 0], [0, 0]]).is_zero_matrix == None
    assert ImmutableMatrix([[x, 0], [0, 0]]).is_zero_matrix == None
    assert ImmutableSparseMatrix([[x, 0], [0, 0]]).is_zero_matrix == None
    assert Matrix([[x, 1], [0, 0]]).is_zero_matrix == False
    a = Symbol('a', nonzero=True)
    assert Matrix([[a, 0], [0, 0]]).is_zero_matrix == False


def test_rotation_matrices():
    # This tests the rotation matrices by rotating about an axis and back.
    theta = pi/3
    r3_plus = rot_axis3(theta)
    r3_minus = rot_axis3(-theta)
    r2_plus = rot_axis2(theta)
    r2_minus = rot_axis2(-theta)
    r1_plus = rot_axis1(theta)
    r1_minus = rot_axis1(-theta)
    assert r3_minus*r3_plus*eye(3) == eye(3)
    assert r2_minus*r2_plus*eye(3) == eye(3)
    assert r1_minus*r1_plus*eye(3) == eye(3)

    # Check the correctness of the trace of the rotation matrix
    assert r1_plus.trace() == 1 + 2*cos(theta)
    assert r2_plus.trace() == 1 + 2*cos(theta)
    assert r3_plus.trace() == 1 + 2*cos(theta)

    # Check that a rotation with zero angle doesn't change anything.
    assert rot_axis1(0) == eye(3)
    assert rot_axis2(0) == eye(3)
    assert rot_axis3(0) == eye(3)

    # Check left-hand convention
    # see Issue #24529
    q1 = Quaternion.from_axis_angle([1, 0, 0], pi / 2)
    q2 = Quaternion.from_axis_angle([0, 1, 0], pi / 2)
    q3 = Quaternion.from_axis_angle([0, 0, 1], pi / 2)
    assert rot_axis1(- pi / 2) == q1.to_rotation_matrix()
    assert rot_axis2(- pi / 2) == q2.to_rotation_matrix()
    assert rot_axis3(- pi / 2) == q3.to_rotation_matrix()
    # Check right-hand convention
    assert rot_ccw_axis1(+ pi / 2) == q1.to_rotation_matrix()
    assert rot_ccw_axis2(+ pi / 2) == q2.to_rotation_matrix()
    assert rot_ccw_axis3(+ pi / 2) == q3.to_rotation_matrix()


def test_DeferredVector():
    assert str(DeferredVector("vector")[4]) == "vector[4]"
    assert sympify(DeferredVector("d")) == DeferredVector("d")
    raises(IndexError, lambda: DeferredVector("d")[-1])
    assert str(DeferredVector("d")) == "d"
    assert repr(DeferredVector("test")) == "DeferredVector('test')"


def test_DeferredVector_not_iterable():
    assert not iterable(DeferredVector('X'))


def test_DeferredVector_Matrix():
    raises(TypeError, lambda: Matrix(DeferredVector("V")))


def test_GramSchmidt():
    R = Rational
    m1 = Matrix(1, 2, [1, 2])
    m2 = Matrix(1, 2, [2, 3])
    assert GramSchmidt([m1, m2]) == \
        [Matrix(1, 2, [1, 2]), Matrix(1, 2, [R(2)/5, R(-1)/5])]
    assert GramSchmidt([m1.T, m2.T]) == \
        [Matrix(2, 1, [1, 2]), Matrix(2, 1, [R(2)/5, R(-1)/5])]
    # from wikipedia
    assert GramSchmidt([Matrix([3, 1]), Matrix([2, 2])], True) == [
        Matrix([3*sqrt(10)/10, sqrt(10)/10]),
        Matrix([-sqrt(10)/10, 3*sqrt(10)/10])]
    # https://github.com/sympy/sympy/issues/9488
    L = FiniteSet(Matrix([1]))
    assert GramSchmidt(L) == [Matrix([[1]])]


def test_casoratian():
    assert casoratian([1, 2, 3, 4], 1) == 0
    assert casoratian([1, 2, 3, 4], 1, zero=False) == 0


def test_zero_dimension_multiply():
    assert (Matrix()*zeros(0, 3)).shape == (0, 3)
    assert zeros(3, 0)*zeros(0, 3) == zeros(3, 3)
    assert zeros(0, 3)*zeros(3, 0) == Matrix()


def test_slice_issue_2884():
    m = Matrix(2, 2, range(4))
    assert m[1, :] == Matrix([[2, 3]])
    assert m[-1, :] == Matrix([[2, 3]])
    assert m[:, 1] == Matrix([[1, 3]]).T
    assert m[:, -1] == Matrix([[1, 3]]).T
    raises(IndexError, lambda: m[2, :])
    raises(IndexError, lambda: m[2, 2])


def test_slice_issue_3401():
    assert zeros(0, 3)[:, -1].shape == (0, 1)
    assert zeros(3, 0)[0, :] == Matrix(1, 0, [])


def test_copyin():
    s = zeros(3, 3)
    s[3] = 1
    assert s[:, 0] == Matrix([0, 1, 0])
    assert s[3] == 1
    assert s[3: 4] == [1]
    s[1, 1] = 42
    assert s[1, 1] == 42
    assert s[1, 1:] == Matrix([[42, 0]])
    s[1, 1:] = Matrix([[5, 6]])
    assert s[1, :] == Matrix([[1, 5, 6]])
    s[1, 1:] = [[42, 43]]
    assert s[1, :] == Matrix([[1, 42, 43]])
    s[0, 0] = 17
    assert s[:, :1] == Matrix([17, 1, 0])
    s[0, 0] = [1, 1, 1]
    assert s[:, 0] == Matrix([1, 1, 1])
    s[0, 0] = Matrix([1, 1, 1])
    assert s[:, 0] == Matrix([1, 1, 1])
    s[0, 0] = SparseMatrix([1, 1, 1])
    assert s[:, 0] == Matrix([1, 1, 1])


def test_invertible_check():
    # sometimes a singular matrix will have a pivot vector shorter than
    # the number of rows in a matrix...
    assert Matrix([[1, 2], [1, 2]]).rref() == (Matrix([[1, 2], [0, 0]]), (0,))
    raises(ValueError, lambda: Matrix([[1, 2], [1, 2]]).inv())
    m = Matrix([
        [-1, -1,  0],
        [ x,  1,  1],
        [ 1,  x, -1],
    ])
    assert len(m.rref()[1]) != m.rows
    # in addition, unless simplify=True in the call to rref, the identity
    # matrix will be returned even though m is not invertible
    assert m.rref()[0] != eye(3)
    assert m.rref(simplify=signsimp)[0] != eye(3)
    raises(ValueError, lambda: m.inv(method="ADJ"))
    raises(ValueError, lambda: m.inv(method="GE"))
    raises(ValueError, lambda: m.inv(method="LU"))


def test_issue_3959():
    x, y = symbols('x, y')
    e = x*y
    assert e.subs(x, Matrix([3, 5, 3])) == Matrix([3, 5, 3])*y


def test_issue_5964():
    assert str(Matrix([[1, 2], [3, 4]])) == 'Matrix([[1, 2], [3, 4]])'


def test_issue_7604():
    x, y = symbols("x y")
    assert sstr(Matrix([[x, 2*y], [y**2, x + 3]])) == \
        'Matrix([\n[   x,   2*y],\n[y**2, x + 3]])'


def test_is_Identity():
    assert eye(3).is_Identity
    assert eye(3).as_immutable().is_Identity
    assert not zeros(3).is_Identity
    assert not ones(3).is_Identity
    # issue 6242
    assert not Matrix([[1, 0, 0]]).is_Identity
    # issue 8854
    assert SparseMatrix(3,3, {(0,0):1, (1,1):1, (2,2):1}).is_Identity
    assert not SparseMatrix(2,3, range(6)).is_Identity
    assert not SparseMatrix(3,3, {(0,0):1, (1,1):1}).is_Identity
    assert not SparseMatrix(3,3, {(0,0):1, (1,1):1, (2,2):1, (0,1):2, (0,2):3}).is_Identity


def test_dot():
    assert ones(1, 3).dot(ones(3, 1)) == 3
    assert ones(1, 3).dot([1, 1, 1]) == 3
    assert Matrix([1, 2, 3]).dot(Matrix([1, 2, 3])) == 14
    assert Matrix([1, 2, 3*I]).dot(Matrix([I, 2, 3*I])) == -5 + I
    assert Matrix([1, 2, 3*I]).dot(Matrix([I, 2, 3*I]), hermitian=False) == -5 + I
    assert Matrix([1, 2, 3*I]).dot(Matrix([I, 2, 3*I]), hermitian=True) == 13 + I
    assert Matrix([1, 2, 3*I]).dot(Matrix([I, 2, 3*I]), hermitian=True, conjugate_convention="physics") == 13 - I
    assert Matrix([1, 2, 3*I]).dot(Matrix([4, 5*I, 6]), hermitian=True, conjugate_convention="right") == 4 + 8*I
    assert Matrix([1, 2, 3*I]).dot(Matrix([4, 5*I, 6]), hermitian=True, conjugate_convention="left") == 4 - 8*I
    assert Matrix([I, 2*I]).dot(Matrix([I, 2*I]), hermitian=False, conjugate_convention="left") == -5
    assert Matrix([I, 2*I]).dot(Matrix([I, 2*I]), conjugate_convention="left") == 5
    raises(ValueError, lambda: Matrix([1, 2]).dot(Matrix([3, 4]), hermitian=True, conjugate_convention="test"))


def test_dual():
    B_x, B_y, B_z, E_x, E_y, E_z = symbols(
        'B_x B_y B_z E_x E_y E_z', real=True)
    F = Matrix((
        (   0,  E_x,  E_y,  E_z),
        (-E_x,    0,  B_z, -B_y),
        (-E_y, -B_z,    0,  B_x),
        (-E_z,  B_y, -B_x,    0)
    ))
    Fd = Matrix((
        (  0, -B_x, -B_y, -B_z),
        (B_x,    0,  E_z, -E_y),
        (B_y, -E_z,    0,  E_x),
        (B_z,  E_y, -E_x,    0)
    ))
    assert F.dual().equals(Fd)
    assert eye(3).dual().equals(zeros(3))
    assert F.dual().dual().equals(-F)


def test_anti_symmetric():
    assert Matrix([1, 2]).is_anti_symmetric() is False
    m = Matrix(3, 3, [0, x**2 + 2*x + 1, y, -(x + 1)**2, 0, x*y, -y, -x*y, 0])
    assert m.is_anti_symmetric() is True
    assert m.is_anti_symmetric(simplify=False) is None
    assert m.is_anti_symmetric(simplify=lambda x: x) is None

    # tweak to fail
    m[2, 1] = -m[2, 1]
    assert m.is_anti_symmetric() is None
    # untweak
    m[2, 1] = -m[2, 1]

    m = m.expand()
    assert m.is_anti_symmetric(simplify=False) is True
    m[0, 0] = 1
    assert m.is_anti_symmetric() is False


def test_normalize_sort_diogonalization():
    A = Matrix(((1, 2), (2, 1)))
    P, Q = A.diagonalize(normalize=True)
    assert P*P.T == P.T*P == eye(P.cols)
    P, Q = A.diagonalize(normalize=True, sort=True)
    assert P*P.T == P.T*P == eye(P.cols)
    assert P*Q*P.inv() == A


def test_issue_5321():
    raises(ValueError, lambda: Matrix([[1, 2, 3], Matrix(0, 1, [])]))


def test_issue_5320():
    assert Matrix.hstack(eye(2), 2*eye(2)) == Matrix([
        [1, 0, 2, 0],
        [0, 1, 0, 2]
    ])
    assert Matrix.vstack(eye(2), 2*eye(2)) == Matrix([
        [1, 0],
        [0, 1],
        [2, 0],
        [0, 2]
    ])
    cls = SparseMatrix
    assert cls.hstack(cls(eye(2)), cls(2*eye(2))) == Matrix([
        [1, 0, 2, 0],
        [0, 1, 0, 2]
    ])


def test_issue_11944():
    A = Matrix([[1]])
    AIm = sympify(A)
    assert Matrix.hstack(AIm, A) == Matrix([[1, 1]])
    assert Matrix.vstack(AIm, A) == Matrix([[1], [1]])


def test_cross():
    a = [1, 2, 3]
    b = [3, 4, 5]
    col = Matrix([-2, 4, -2])
    row = col.T

    def test(M, ans):
        assert ans == M
        assert type(M) == cls
    for cls in all_classes:
        A = cls(a)
        B = cls(b)
        test(A.cross(B), col)
        test(A.cross(B.T), col)
        test(A.T.cross(B.T), row)
        test(A.T.cross(B), row)
    raises(ShapeError, lambda:
        Matrix(1, 2, [1, 1]).cross(Matrix(1, 2, [1, 1])))


def test_hat_vee():
    v1 = Matrix([x, y, z])
    v2 = Matrix([a, b, c])
    assert v1.hat() * v2 == v1.cross(v2)
    assert v1.hat().is_anti_symmetric()
    assert v1.hat().vee() == v1


def test_hash():
    for cls in immutable_classes:
        s = {cls.eye(1), cls.eye(1)}
        assert len(s) == 1 and s.pop() == cls.eye(1)
    # issue 3979
    for cls in mutable_classes:
        assert not isinstance(cls.eye(1), Hashable)


def test_adjoint():
    dat = [[0, I], [1, 0]]
    ans = Matrix([[0, 1], [-I, 0]])
    for cls in all_classes:
        assert ans == cls(dat).adjoint()


def test_atoms():
    m = Matrix([[1, 2], [x, 1 - 1/x]])
    assert m.atoms() == {S.One,S(2),S.NegativeOne, x}
    assert m.atoms(Symbol) == {x}


def test_pinv():
    # Pseudoinverse of an invertible matrix is the inverse.
    A1 = Matrix([[a, b], [c, d]])
    assert simplify(A1.pinv(method="RD")) == simplify(A1.inv())

    # Test the four properties of the pseudoinverse for various matrices.
    As = [Matrix([[13, 104], [2212, 3], [-3, 5]]),
          Matrix([[1, 7, 9], [11, 17, 19]]),
          Matrix([a, b])]

    for A in As:
        A_pinv = A.pinv(method="RD")
        AAp = A * A_pinv
        ApA = A_pinv * A
        assert simplify(AAp * A) == A
        assert simplify(ApA * A_pinv) == A_pinv
        assert AAp.H == AAp
        assert ApA.H == ApA

    # XXX Pinv with diagonalization makes expression too complicated.
    for A in As:
        A_pinv = simplify(A.pinv(method="ED"))
        AAp = A * A_pinv
        ApA = A_pinv * A
        assert simplify(AAp * A) == A
        assert simplify(ApA * A_pinv) == A_pinv
        assert AAp.H == AAp
        assert ApA.H == ApA

    # XXX Computing pinv using diagonalization makes an expression that
    # is too complicated to simplify.
    # A1 = Matrix([[a, b], [c, d]])
    # assert simplify(A1.pinv(method="ED")) == simplify(A1.inv())
    # so this is tested numerically at a fixed random point

    from sympy.core.numbers import comp
    q = A1.pinv(method="ED")
    w = A1.inv()
    reps = {a: -73633, b: 11362, c: 55486, d: 62570}
    assert all(
        comp(i.n(), j.n())
        for i, j in zip(q.subs(reps), w.subs(reps))
        )


@slow
def test_pinv_rank_deficient_when_diagonalization_fails():
    # Test the four properties of the pseudoinverse for matrices when
    # diagonalization of A.H*A fails.
    As = [
        Matrix([
            [61, 89, 55, 20, 71, 0],
            [62, 96, 85, 85, 16, 0],
            [69, 56, 17,  4, 54, 0],
            [10, 54, 91, 41, 71, 0],
            [ 7, 30, 10, 48, 90, 0],
            [0, 0, 0, 0, 0, 0]])
    ]
    for A in As:
        A_pinv = A.pinv(method="ED")
        AAp = A * A_pinv
        ApA = A_pinv * A
        assert AAp.H == AAp

        # Here ApA.H and ApA are equivalent expressions but they are very
        # complicated expressions involving RootOfs. Using simplify would be
        # too slow and so would evalf so we substitute approximate values for
        # the RootOfs and then evalf which is less accurate but good enough to
        # confirm that these two matrices are equivalent.
        #
        # assert ApA.H == ApA  # <--- would fail (structural equality)
        # assert simplify(ApA.H - ApA).is_zero_matrix  # <--- too slow
        # (ApA.H - ApA).evalf()  # <--- too slow

        def allclose(M1, M2):
            rootofs = M1.atoms(RootOf)
            rootofs_approx = {r: r.evalf() for r in rootofs}
            diff_approx = (M1 - M2).xreplace(rootofs_approx).evalf()
            return all(abs(e) < 1e-10 for e in diff_approx)

        assert allclose(ApA.H, ApA)


def test_issue_7201():
    assert ones(0, 1) + ones(0, 1) == Matrix(0, 1, [])
    assert ones(1, 0) + ones(1, 0) == Matrix(1, 0, [])


def test_free_symbols():
    for M in ImmutableMatrix, ImmutableSparseMatrix, Matrix, SparseMatrix:
        assert M([[x], [0]]).free_symbols == {x}


def test_from_ndarray():
    """See issue 7465."""
    try:
        from numpy import array
    except ImportError:
        skip('NumPy must be available to test creating matrices from ndarrays')

    assert Matrix(array([1, 2, 3])) == Matrix([1, 2, 3])
    assert Matrix(array([[1, 2, 3]])) == Matrix([[1, 2, 3]])
    assert Matrix(array([[1, 2, 3], [4, 5, 6]])) == \
        Matrix([[1, 2, 3], [4, 5, 6]])
    assert Matrix(array([x, y, z])) == Matrix([x, y, z])
    raises(NotImplementedError,
        lambda: Matrix(array([[[1, 2], [3, 4]], [[5, 6], [7, 8]]])))
    assert Matrix([array([1, 2]), array([3, 4])]) == Matrix([[1, 2], [3, 4]])
    assert Matrix([array([1, 2]), [3, 4]]) == Matrix([[1, 2], [3, 4]])
    assert Matrix([array([]), array([])]) == Matrix(2, 0, []) != Matrix([])


def test_17522_numpy():
    from sympy.matrices.common import _matrixify
    try:
        from numpy import array, matrix
    except ImportError:
        skip('NumPy must be available to test indexing matrixified NumPy ndarrays and matrices')

    m = _matrixify(array([[1, 2], [3, 4]]))
    assert m[3] == 4
    assert list(m) == [1, 2, 3, 4]

    with ignore_warnings(PendingDeprecationWarning):
        m = _matrixify(matrix([[1, 2], [3, 4]]))
    assert m[3] == 4
    assert list(m) == [1, 2, 3, 4]


def test_17522_mpmath():
    from sympy.matrices.common import _matrixify
    try:
        from mpmath import matrix
    except ImportError:
        skip('mpmath must be available to test indexing matrixified mpmath matrices')

    m = _matrixify(matrix([[1, 2], [3, 4]]))
    assert m[3] == 4.0
    assert list(m) == [1.0, 2.0, 3.0, 4.0]


def test_17522_scipy():
    from sympy.matrices.common import _matrixify
    try:
        from scipy.sparse import csr_matrix
    except ImportError:
        skip('SciPy must be available to test indexing matrixified SciPy sparse matrices')

    m = _matrixify(csr_matrix([[1, 2], [3, 4]]))
    assert m[3] == 4
    assert list(m) == [1, 2, 3, 4]


def test_hermitian():
    a = Matrix([[1, I], [-I, 1]])
    assert a.is_hermitian
    a[0, 0] = 2*I
    assert a.is_hermitian is False
    a[0, 0] = x
    assert a.is_hermitian is None
    a[0, 1] = a[1, 0]*I
    assert a.is_hermitian is False


def test_issue_9457_9467_9876():
    # for row_del(index)
    M = Matrix([[1, 2, 3], [2, 3, 4], [3, 4, 5]])
    M.row_del(1)
    assert M == Matrix([[1, 2, 3], [3, 4, 5]])
    N = Matrix([[1, 2, 3], [2, 3, 4], [3, 4, 5]])
    N.row_del(-2)
    assert N == Matrix([[1, 2, 3], [3, 4, 5]])
    O = Matrix([[1, 2, 3], [5, 6, 7], [9, 10, 11]])
    O.row_del(-1)
    assert O == Matrix([[1, 2, 3], [5, 6, 7]])
    P = Matrix([[1, 2, 3], [2, 3, 4], [3, 4, 5]])
    raises(IndexError, lambda: P.row_del(10))
    Q = Matrix([[1, 2, 3], [2, 3, 4], [3, 4, 5]])
    raises(IndexError, lambda: Q.row_del(-10))

    # for col_del(index)
    M = Matrix([[1, 2, 3], [2, 3, 4], [3, 4, 5]])
    M.col_del(1)
    assert M == Matrix([[1, 3], [2, 4], [3, 5]])
    N = Matrix([[1, 2, 3], [2, 3, 4], [3, 4, 5]])
    N.col_del(-2)
    assert N == Matrix([[1, 3], [2, 4], [3, 5]])
    P = Matrix([[1, 2, 3], [2, 3, 4], [3, 4, 5]])
    raises(IndexError, lambda: P.col_del(10))
    Q = Matrix([[1, 2, 3], [2, 3, 4], [3, 4, 5]])
    raises(IndexError, lambda: Q.col_del(-10))


def test_issue_9422():
    x, y = symbols('x y', commutative=False)
    a, b = symbols('a b')
    M = eye(2)
    M1 = Matrix(2, 2, [x, y, y, z])
    assert y*x*M != x*y*M
    assert b*a*M == a*b*M
    assert x*M1 != M1*x
    assert a*M1 == M1*a
    assert y*x*M == Matrix([[y*x, 0], [0, y*x]])


def test_issue_10770():
    M = Matrix([])
    a = ['col_insert', 'row_join'], Matrix([9, 6, 3])
    b = ['row_insert', 'col_join'], a[1].T
    c = ['row_insert', 'col_insert'], Matrix([[1, 2], [3, 4]])
    for ops, m in (a, b, c):
        for op in ops:
            f = getattr(M, op)
            new = f(m) if 'join' in op else f(42, m)
            assert new == m and id(new) != id(m)


def test_issue_10658():
    A = Matrix([[1, 2, 3], [4, 5, 6], [7, 8, 9]])
    assert A.extract([0, 1, 2], [True, True, False]) == \
        Matrix([[1, 2], [4, 5], [7, 8]])
    assert A.extract([0, 1, 2], [True, False, False]) == Matrix([[1], [4], [7]])
    assert A.extract([True, False, False], [0, 1, 2]) == Matrix([[1, 2, 3]])
    assert A.extract([True, False, True], [0, 1, 2]) == \
        Matrix([[1, 2, 3], [7, 8, 9]])
    assert A.extract([0, 1, 2], [False, False, False]) == Matrix(3, 0, [])
    assert A.extract([False, False, False], [0, 1, 2]) == Matrix(0, 3, [])
    assert A.extract([True, False, True], [False, True, False]) == \
        Matrix([[2], [8]])


def test_opportunistic_simplification():
    # this test relates to issue #10718, #9480, #11434

    # issue #9480
    m = Matrix([[-5 + 5*sqrt(2), -5], [-5*sqrt(2)/2 + 5, -5*sqrt(2)/2]])
    assert m.rank() == 1

    # issue #10781
    m = Matrix([[3+3*sqrt(3)*I, -9],[4,-3+3*sqrt(3)*I]])
    assert simplify(m.rref()[0] - Matrix([[1, -9/(3 + 3*sqrt(3)*I)], [0, 0]])) == zeros(2, 2)

    # issue #11434
    ax,ay,bx,by,cx,cy,dx,dy,ex,ey,t0,t1 = symbols('a_x a_y b_x b_y c_x c_y d_x d_y e_x e_y t_0 t_1')
    m = Matrix([[ax,ay,ax*t0,ay*t0,0],[bx,by,bx*t0,by*t0,0],[cx,cy,cx*t0,cy*t0,1],[dx,dy,dx*t0,dy*t0,1],[ex,ey,2*ex*t1-ex*t0,2*ey*t1-ey*t0,0]])
    assert m.rank() == 4


def test_partial_pivoting():
    # example from https://en.wikipedia.org/wiki/Pivot_element
    # partial pivoting with back substitution gives a perfect result
    # naive pivoting give an error ~1e-13, so anything better than
    # 1e-15 is good
    mm=Matrix([[0.003, 59.14, 59.17], [5.291, -6.13, 46.78]])
    assert (mm.rref()[0] - Matrix([[1.0,   0, 10.0],
                                   [  0, 1.0,  1.0]])).norm() < 1e-15

    # issue #11549
    m_mixed = Matrix([[6e-17, 1.0, 4],
                      [ -1.0,   0, 8],
                      [    0,   0, 1]])
    m_float = Matrix([[6e-17,  1.0, 4.],
                      [ -1.0,   0., 8.],
                      [   0.,   0., 1.]])
    m_inv = Matrix([[  0,    -1.0,  8.0],
                    [1.0, 6.0e-17, -4.0],
                    [  0,       0,    1]])
    # this example is numerically unstable and involves a matrix with a norm >= 8,
    # this comparing the difference of the results with 1e-15 is numerically sound.
    assert (m_mixed.inv() - m_inv).norm() < 1e-15
    assert (m_float.inv() - m_inv).norm() < 1e-15


def test_iszero_substitution():
    """ When doing numerical computations, all elements that pass
    the iszerofunc test should be set to numerically zero if they
    aren't already. """

    # Matrix from issue #9060
    m = Matrix([[0.9, -0.1, -0.2, 0],[-0.8, 0.9, -0.4, 0],[-0.1, -0.8, 0.6, 0]])
    m_rref = m.rref(iszerofunc=lambda x: abs(x)<6e-15)[0]
    m_correct = Matrix([[1.0,   0, -0.301369863013699, 0],[  0, 1.0, -0.712328767123288, 0],[  0,   0,                  0, 0]])
    m_diff = m_rref - m_correct
    assert m_diff.norm() < 1e-15
    # if a zero-substitution wasn't made, this entry will be -1.11022302462516e-16
    assert m_rref[2,2] == 0


def test_issue_11238():
    from sympy.geometry.point import Point
    xx = 8*tan(pi*Rational(13, 45))/(tan(pi*Rational(13, 45)) + sqrt(3))
    yy = (-8*sqrt(3)*tan(pi*Rational(13, 45))**2 + 24*tan(pi*Rational(13, 45)))/(-3 + tan(pi*Rational(13, 45))**2)
    p1 = Point(0, 0)
    p2 = Point(1, -sqrt(3))
    p0 = Point(xx,yy)
    m1 = Matrix([p1 - simplify(p0), p2 - simplify(p0)])
    m2 = Matrix([p1 - p0, p2 - p0])
    m3 = Matrix([simplify(p1 - p0), simplify(p2 - p0)])

    # This system has expressions which are zero and
    # cannot be easily proved to be such, so without
    # numerical testing, these assertions will fail.
    Z = lambda x: abs(x.n()) < 1e-20
    assert m1.rank(simplify=True, iszerofunc=Z) == 1
    assert m2.rank(simplify=True, iszerofunc=Z) == 1
    assert m3.rank(simplify=True, iszerofunc=Z) == 1


def test_as_real_imag():
    m1 = Matrix(2,2,[1,2,3,4])
    m2 = m1*S.ImaginaryUnit
    m3 = m1 + m2

    for kls in all_classes:
        a,b = kls(m3).as_real_imag()
        assert list(a) == list(m1)
        assert list(b) == list(m1)


def test_deprecated():
    # Maintain tests for deprecated functions.  We must capture
    # the deprecation warnings.  When the deprecated functionality is
    # removed, the corresponding tests should be removed.

    m = Matrix(3, 3, [0, 1, 0, -4, 4, 0, -2, 1, 2])
    P, Jcells = m.jordan_cells()
    assert Jcells[1] == Matrix(1, 1, [2])
    assert Jcells[0] == Matrix(2, 2, [2, 1, 0, 2])


def test_issue_14489():
    from sympy.core.mod import Mod
    A = Matrix([-1, 1, 2])
    B = Matrix([10, 20, -15])

    assert Mod(A, 3) == Matrix([2, 1, 2])
    assert Mod(B, 4) == Matrix([2, 0, 1])


def test_issue_14943():
    # Test that __array__ accepts the optional dtype argument
    try:
        from numpy import array
    except ImportError:
        skip('NumPy must be available to test creating matrices from ndarrays')

    M = Matrix([[1,2], [3,4]])
    assert array(M, dtype=float).dtype.name == 'float64'


def test_case_6913():
    m = MatrixSymbol('m', 1, 1)
    a = Symbol("a")
    a = m[0, 0]>0
    assert str(a) == 'm[0, 0] > 0'


def test_issue_11948():
    A = MatrixSymbol('A', 3, 3)
    a = Wild('a')
    assert A.match(a) == {a: A}


def test_gramschmidt_conjugate_dot():
    vecs = [Matrix([1, I]), Matrix([1, -I])]
    assert Matrix.orthogonalize(*vecs) == \
        [Matrix([[1], [I]]), Matrix([[1], [-I]])]

    vecs = [Matrix([1, I, 0]), Matrix([I, 0, -I])]
    assert Matrix.orthogonalize(*vecs) == \
        [Matrix([[1], [I], [0]]), Matrix([[I/2], [S(1)/2], [-I]])]

    mat = Matrix([[1, I], [1, -I]])
    Q, R = mat.QRdecomposition()
    assert Q * Q.H == Matrix.eye(2)


def test_issue_8207():
    a = Matrix(MatrixSymbol('a', 3, 1))
    b = Matrix(MatrixSymbol('b', 3, 1))
    c = a.dot(b)
    d = diff(c, a[0, 0])
    e = diff(d, a[0, 0])
    assert d == b[0, 0]
    assert e == 0


def test_func():
    from sympy.simplify.simplify import nthroot

    A = Matrix([[1, 2],[0, 3]])
    assert A.analytic_func(sin(x*t), x) == Matrix([[sin(t), sin(3*t) - sin(t)], [0, sin(3*t)]])

    A = Matrix([[2, 1],[1, 2]])
    assert (pi * A / 6).analytic_func(cos(x), x) == Matrix([[sqrt(3)/4, -sqrt(3)/4], [-sqrt(3)/4, sqrt(3)/4]])


    raises(ValueError, lambda : zeros(5).analytic_func(log(x), x))
    raises(ValueError, lambda : (A*x).analytic_func(log(x), x))

    A = Matrix([[0, -1, -2, 3], [0, -1, -2, 3], [0, 1, 0, -1], [0, 0, -1, 1]])
    assert A.analytic_func(exp(x), x) == A.exp()
    raises(ValueError, lambda : A.analytic_func(sqrt(x), x))

    A = Matrix([[41, 12],[12, 34]])
    assert simplify(A.analytic_func(sqrt(x), x)**2) == A

    A = Matrix([[3, -12, 4], [-1, 0, -2], [-1, 5, -1]])
    assert simplify(A.analytic_func(nthroot(x, 3), x)**3) == A

    A = Matrix([[2, 0, 0, 0], [1, 2, 0, 0], [0, 1, 3, 0], [0, 0, 1, 3]])
    assert A.analytic_func(exp(x), x) == A.exp()

    A = Matrix([[0, 2, 1, 6], [0, 0, 1, 2], [0, 0, 0, 3], [0, 0, 0, 0]])
    assert A.analytic_func(exp(x*t), x) == expand(simplify((A*t).exp()))


@skip_under_pyodide("Cannot create threads under pyodide.")
def test_issue_19809():

    def f():
        assert _dotprodsimp_state.state == None
        m = Matrix([[1]])
        m = m * m
        return True

    with dotprodsimp(True):
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future = executor.submit(f)
            assert future.result()


def test_issue_23276():
    M = Matrix([x, y])
    assert integrate(M, (x, 0, 1), (y, 0, 1)) == Matrix([
        [S.Half],
        [S.Half]])


def test_issue_27225():
    # https://github.com/sympy/sympy/issues/27225
    raises(TypeError, lambda : floor(Matrix([1, 1, 0])))