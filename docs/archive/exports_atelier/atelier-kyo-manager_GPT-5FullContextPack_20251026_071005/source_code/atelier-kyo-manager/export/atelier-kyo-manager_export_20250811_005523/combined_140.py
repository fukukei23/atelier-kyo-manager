
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_core\_tests\test_asyncgen.py ===
from __future__ import annotations

import contextlib
import sys
import weakref
from math import inf
from types import AsyncGeneratorType
from typing import TYPE_CHECKING, NoReturn

import pytest

from ... import _core
from .tutil import gc_collect_harder, restore_unraisablehook

if TYPE_CHECKING:
    from collections.abc import AsyncGenerator


def test_asyncgen_basics() -> None:
    collected = []

    async def example(cause: str) -> AsyncGenerator[int, None]:
        try:
            with contextlib.suppress(GeneratorExit):
                yield 42
            await _core.checkpoint()
        except _core.Cancelled:
            assert "exhausted" not in cause
            task_name = _core.current_task().name
            assert cause in task_name or task_name == "<init>"
            assert _core.current_effective_deadline() == -inf
            with pytest.raises(_core.Cancelled):
                await _core.checkpoint()
            collected.append(cause)
        else:
            assert "async_main" in _core.current_task().name
            assert "exhausted" in cause
            assert _core.current_effective_deadline() == inf
            await _core.checkpoint()
            collected.append(cause)

    saved = []

    async def async_main() -> None:
        # GC'ed before exhausted
        with pytest.warns(  # noqa: PT031
            ResourceWarning,
            match="Async generator.*collected before.*exhausted",
        ):
            assert await example("abandoned").asend(None) == 42
            gc_collect_harder()
        await _core.wait_all_tasks_blocked()
        assert collected.pop() == "abandoned"

        aiter_ = example("exhausted 1")
        try:
            assert await aiter_.asend(None) == 42
        finally:
            await aiter_.aclose()
        assert collected.pop() == "exhausted 1"

        # Also fine if you exhaust it at point of use
        async for val in example("exhausted 2"):
            assert val == 42
        assert collected.pop() == "exhausted 2"

        gc_collect_harder()

        # No problems saving the geniter when using either of these patterns
        aiter_ = example("exhausted 3")
        try:
            saved.append(aiter_)
            assert await aiter_.asend(None) == 42
        finally:
            await aiter_.aclose()
        assert collected.pop() == "exhausted 3"

        # Also fine if you exhaust it at point of use
        saved.append(example("exhausted 4"))
        async for val in saved[-1]:
            assert val == 42
        assert collected.pop() == "exhausted 4"

        # Leave one referenced-but-unexhausted and make sure it gets cleaned up
        saved.append(example("outlived run"))
        assert await saved[-1].asend(None) == 42
        assert collected == []

    _core.run(async_main)
    assert collected.pop() == "outlived run"
    for agen in saved:
        assert isinstance(agen, AsyncGeneratorType)
        assert agen.ag_frame is None  # all should now be exhausted


async def test_asyncgen_throws_during_finalization(
    caplog: pytest.LogCaptureFixture,
) -> None:
    record = []

    async def agen() -> AsyncGenerator[int, None]:
        try:
            yield 1
        finally:
            await _core.cancel_shielded_checkpoint()
            record.append("crashing")
            raise ValueError("oops")

    with restore_unraisablehook():
        await agen().asend(None)
        gc_collect_harder()
    await _core.wait_all_tasks_blocked()
    assert record == ["crashing"]
    # Following type ignore is because typing for LogCaptureFixture is wrong
    exc_type, exc_value, _exc_traceback = caplog.records[0].exc_info  # type: ignore[misc]
    assert exc_type is ValueError
    assert str(exc_value) == "oops"
    assert "during finalization of async generator" in caplog.records[0].message


def test_firstiter_after_closing() -> None:
    saved = []
    record = []

    async def funky_agen() -> AsyncGenerator[int, None]:
        try:
            yield 1
        except GeneratorExit:
            record.append("cleanup 1")
            raise
        try:
            yield 2
        finally:
            record.append("cleanup 2")
            await funky_agen().asend(None)

    async def async_main() -> None:
        aiter_ = funky_agen()
        saved.append(aiter_)
        assert await aiter_.asend(None) == 1
        assert await aiter_.asend(None) == 2

    _core.run(async_main)
    assert record == ["cleanup 2", "cleanup 1"]


def test_interdependent_asyncgen_cleanup_order() -> None:
    saved: list[AsyncGenerator[int, None]] = []
    record: list[int | str] = []

    async def innermost() -> AsyncGenerator[int, None]:
        try:
            yield 1
        finally:
            await _core.cancel_shielded_checkpoint()
            record.append("innermost")

    async def agen(
        label: int,
        inner: AsyncGenerator[int, None],
    ) -> AsyncGenerator[int, None]:
        try:
            yield await inner.asend(None)
        finally:
            # Either `inner` has already been cleaned up, or
            # we're about to exhaust it. Either way, we wind
            # up with `record` containing the labels in
            # innermost-to-outermost order.
            with pytest.raises(StopAsyncIteration):
                await inner.asend(None)
            record.append(label)

    async def async_main() -> None:
        # This makes a chain of 101 interdependent asyncgens:
        # agen(99)'s cleanup will iterate agen(98)'s will iterate
        # ... agen(0)'s will iterate innermost()'s
        ag_chain = innermost()
        for idx in range(100):
            ag_chain = agen(idx, ag_chain)
        saved.append(ag_chain)
        assert await ag_chain.asend(None) == 1
        assert record == []

    _core.run(async_main)
    assert record == ["innermost", *range(100)]


@restore_unraisablehook()
def test_last_minute_gc_edge_case() -> None:
    saved: list[AsyncGenerator[int, None]] = []
    record = []
    needs_retry = True

    async def agen() -> AsyncGenerator[int, None]:
        try:
            yield 1
        finally:
            record.append("cleaned up")

    def collect_at_opportune_moment(token: _core._entry_queue.TrioToken) -> None:
        runner = _core._run.GLOBAL_RUN_CONTEXT.runner
        assert runner.system_nursery is not None
        if runner.system_nursery._closed and isinstance(
            runner.asyncgens.alive,
            weakref.WeakSet,
        ):
            saved.clear()
            record.append("final collection")
            gc_collect_harder()
            record.append("done")
        else:
            try:
                token.run_sync_soon(collect_at_opportune_moment, token)
            except _core.RunFinishedError:  # pragma: no cover
                nonlocal needs_retry
                needs_retry = True

    async def async_main() -> None:
        token = _core.current_trio_token()
        token.run_sync_soon(collect_at_opportune_moment, token)
        saved.append(agen())
        await saved[-1].asend(None)

    # Actually running into the edge case requires that the run_sync_soon task
    # execute in between the system nursery's closure and the strong-ification
    # of runner.asyncgens. There's about a 25% chance that it doesn't
    # (if the run_sync_soon task runs before init on one tick and after init
    # on the next tick); if we try enough times, we can make the chance of
    # failure as small as we want.
    for _attempt in range(50):
        needs_retry = False
        record.clear()
        saved.clear()
        _core.run(async_main)
        if needs_retry:  # pragma: no cover
            assert record == ["cleaned up"]
        else:
            assert record == ["final collection", "done", "cleaned up"]
            break
    else:  # pragma: no cover
        pytest.fail(
            "Didn't manage to hit the trailing_finalizer_asyncgens case "
            f"despite trying {_attempt} times",
        )


async def step_outside_async_context(aiter_: AsyncGenerator[int, None]) -> None:
    # abort_fns run outside of task context, at least if they're
    # triggered by a deadline expiry rather than a direct
    # cancellation.  Thus, an asyncgen first iterated inside one
    # will appear non-Trio, and since no other hooks were installed,
    # will use the last-ditch fallback handling (that tries to mimic
    # CPython's behavior with no hooks).
    #
    # NB: the strangeness with aiter being an attribute of abort_fn is
    # to make it as easy as possible to ensure we don't hang onto a
    # reference to aiter inside the guts of the run loop.
    def abort_fn(_: _core.RaiseCancelT) -> _core.Abort:
        with pytest.raises(StopIteration, match="42"):
            abort_fn.aiter.asend(None).send(None)  # type: ignore[attr-defined]  # Callables don't have attribute "aiter"
        del abort_fn.aiter  # type: ignore[attr-defined]
        return _core.Abort.SUCCEEDED

    abort_fn.aiter = aiter_  # type: ignore[attr-defined]

    async with _core.open_nursery() as nursery:
        nursery.start_soon(_core.wait_task_rescheduled, abort_fn)
        await _core.wait_all_tasks_blocked()
        nursery.cancel_scope.deadline = _core.current_time()


async def test_fallback_when_no_hook_claims_it(
    capsys: pytest.CaptureFixture[str],
) -> None:
    async def well_behaved() -> AsyncGenerator[int, None]:
        yield 42

    async def yields_after_yield() -> AsyncGenerator[int, None]:
        with pytest.raises(GeneratorExit):
            yield 42
        yield 100

    async def awaits_after_yield() -> AsyncGenerator[int, None]:
        with pytest.raises(GeneratorExit):
            yield 42
        await _core.cancel_shielded_checkpoint()

    with restore_unraisablehook():
        await step_outside_async_context(well_behaved())
        gc_collect_harder()
        assert capsys.readouterr().err == ""

        await step_outside_async_context(yields_after_yield())
        gc_collect_harder()
        assert "ignored GeneratorExit" in capsys.readouterr().err

        await step_outside_async_context(awaits_after_yield())
        gc_collect_harder()
        assert "awaited something during finalization" in capsys.readouterr().err


def test_delegation_to_existing_hooks() -> None:
    record = []

    def my_firstiter(agen: AsyncGenerator[object, NoReturn]) -> None:
        assert isinstance(agen, AsyncGeneratorType)
        record.append("firstiter " + agen.ag_frame.f_locals["arg"])

    def my_finalizer(agen: AsyncGenerator[object, NoReturn]) -> None:
        assert isinstance(agen, AsyncGeneratorType)
        record.append("finalizer " + agen.ag_frame.f_locals["arg"])

    async def example(arg: str) -> AsyncGenerator[int, None]:
        try:
            yield 42
        finally:
            with pytest.raises(_core.Cancelled):
                await _core.checkpoint()
            record.append("trio collected " + arg)

    async def async_main() -> None:
        await step_outside_async_context(example("theirs"))
        assert await example("ours").asend(None) == 42
        gc_collect_harder()
        assert record == ["firstiter theirs", "finalizer theirs"]
        record[:] = []
        await _core.wait_all_tasks_blocked()
        assert record == ["trio collected ours"]

    with restore_unraisablehook():
        old_hooks = sys.get_asyncgen_hooks()
        sys.set_asyncgen_hooks(my_firstiter, my_finalizer)
        try:
            _core.run(async_main)
        finally:
            assert sys.get_asyncgen_hooks() == (my_firstiter, my_finalizer)
            sys.set_asyncgen_hooks(*old_hooks)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pycparser\_ast_gen.py ===
#-----------------------------------------------------------------
# _ast_gen.py
#
# Generates the AST Node classes from a specification given in
# a configuration file
#
# The design of this module was inspired by astgen.py from the
# Python 2.5 code-base.
#
# Eli Bendersky [https://eli.thegreenplace.net/]
# License: BSD
#-----------------------------------------------------------------
from string import Template


class ASTCodeGenerator(object):
    def __init__(self, cfg_filename='_c_ast.cfg'):
        """ Initialize the code generator from a configuration
            file.
        """
        self.cfg_filename = cfg_filename
        self.node_cfg = [NodeCfg(name, contents)
            for (name, contents) in self.parse_cfgfile(cfg_filename)]

    def generate(self, file=None):
        """ Generates the code into file, an open file buffer.
        """
        src = Template(_PROLOGUE_COMMENT).substitute(
            cfg_filename=self.cfg_filename)

        src += _PROLOGUE_CODE
        for node_cfg in self.node_cfg:
            src += node_cfg.generate_source() + '\n\n'

        file.write(src)

    def parse_cfgfile(self, filename):
        """ Parse the configuration file and yield pairs of
            (name, contents) for each node.
        """
        with open(filename, "r") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                colon_i = line.find(':')
                lbracket_i = line.find('[')
                rbracket_i = line.find(']')
                if colon_i < 1 or lbracket_i <= colon_i or rbracket_i <= lbracket_i:
                    raise RuntimeError("Invalid line in %s:\n%s\n" % (filename, line))

                name = line[:colon_i]
                val = line[lbracket_i + 1:rbracket_i]
                vallist = [v.strip() for v in val.split(',')] if val else []
                yield name, vallist


class NodeCfg(object):
    """ Node configuration.

        name: node name
        contents: a list of contents - attributes and child nodes
        See comment at the top of the configuration file for details.
    """

    def __init__(self, name, contents):
        self.name = name
        self.all_entries = []
        self.attr = []
        self.child = []
        self.seq_child = []

        for entry in contents:
            clean_entry = entry.rstrip('*')
            self.all_entries.append(clean_entry)

            if entry.endswith('**'):
                self.seq_child.append(clean_entry)
            elif entry.endswith('*'):
                self.child.append(clean_entry)
            else:
                self.attr.append(entry)

    def generate_source(self):
        src = self._gen_init()
        src += '\n' + self._gen_children()
        src += '\n' + self._gen_iter()
        src += '\n' + self._gen_attr_names()
        return src

    def _gen_init(self):
        src = "class %s(Node):\n" % self.name

        if self.all_entries:
            args = ', '.join(self.all_entries)
            slots = ', '.join("'{0}'".format(e) for e in self.all_entries)
            slots += ", 'coord', '__weakref__'"
            arglist = '(self, %s, coord=None)' % args
        else:
            slots = "'coord', '__weakref__'"
            arglist = '(self, coord=None)'

        src += "    __slots__ = (%s)\n" % slots
        src += "    def __init__%s:\n" % arglist

        for name in self.all_entries + ['coord']:
            src += "        self.%s = %s\n" % (name, name)

        return src

    def _gen_children(self):
        src = '    def children(self):\n'

        if self.all_entries:
            src += '        nodelist = []\n'

            for child in self.child:
                src += (
                    '        if self.%(child)s is not None:' +
                    ' nodelist.append(("%(child)s", self.%(child)s))\n') % (
                        dict(child=child))

            for seq_child in self.seq_child:
                src += (
                    '        for i, child in enumerate(self.%(child)s or []):\n'
                    '            nodelist.append(("%(child)s[%%d]" %% i, child))\n') % (
                        dict(child=seq_child))

            src += '        return tuple(nodelist)\n'
        else:
            src += '        return ()\n'

        return src

    def _gen_iter(self):
        src = '    def __iter__(self):\n'

        if self.all_entries:
            for child in self.child:
                src += (
                    '        if self.%(child)s is not None:\n' +
                    '            yield self.%(child)s\n') % (dict(child=child))

            for seq_child in self.seq_child:
                src += (
                    '        for child in (self.%(child)s or []):\n'
                    '            yield child\n') % (dict(child=seq_child))

            if not (self.child or self.seq_child):
                # Empty generator
                src += (
                    '        return\n' +
                    '        yield\n')
        else:
            # Empty generator
            src += (
                '        return\n' +
                '        yield\n')

        return src

    def _gen_attr_names(self):
        src = "    attr_names = (" + ''.join("%r, " % nm for nm in self.attr) + ')'
        return src


_PROLOGUE_COMMENT = \
r'''#-----------------------------------------------------------------
# ** ATTENTION **
# This code was automatically generated from the file:
# $cfg_filename
#
# Do not modify it directly. Modify the configuration file and
# run the generator again.
# ** ** *** ** **
#
# pycparser: c_ast.py
#
# AST Node classes.
#
# Eli Bendersky [https://eli.thegreenplace.net/]
# License: BSD
#-----------------------------------------------------------------

'''

_PROLOGUE_CODE = r'''
import sys

def _repr(obj):
    """
    Get the representation of an object, with dedicated pprint-like format for lists.
    """
    if isinstance(obj, list):
        return '[' + (',\n '.join((_repr(e).replace('\n', '\n ') for e in obj))) + '\n]'
    else:
        return repr(obj)

class Node(object):
    __slots__ = ()
    """ Abstract base class for AST nodes.
    """
    def __repr__(self):
        """ Generates a python representation of the current node
        """
        result = self.__class__.__name__ + '('

        indent = ''
        separator = ''
        for name in self.__slots__[:-2]:
            result += separator
            result += indent
            result += name + '=' + (_repr(getattr(self, name)).replace('\n', '\n  ' + (' ' * (len(name) + len(self.__class__.__name__)))))

            separator = ','
            indent = '\n ' + (' ' * len(self.__class__.__name__))

        result += indent + ')'

        return result

    def children(self):
        """ A sequence of all children that are Nodes
        """
        pass

    def show(self, buf=sys.stdout, offset=0, attrnames=False, nodenames=False, showcoord=False, _my_node_name=None):
        """ Pretty print the Node and all its attributes and
            children (recursively) to a buffer.

            buf:
                Open IO buffer into which the Node is printed.

            offset:
                Initial offset (amount of leading spaces)

            attrnames:
                True if you want to see the attribute names in
                name=value pairs. False to only see the values.

            nodenames:
                True if you want to see the actual node names
                within their parents.

            showcoord:
                Do you want the coordinates of each Node to be
                displayed.
        """
        lead = ' ' * offset
        if nodenames and _my_node_name is not None:
            buf.write(lead + self.__class__.__name__+ ' <' + _my_node_name + '>: ')
        else:
            buf.write(lead + self.__class__.__name__+ ': ')

        if self.attr_names:
            if attrnames:
                nvlist = [(n, getattr(self,n)) for n in self.attr_names]
                attrstr = ', '.join('%s=%s' % nv for nv in nvlist)
            else:
                vlist = [getattr(self, n) for n in self.attr_names]
                attrstr = ', '.join('%s' % v for v in vlist)
            buf.write(attrstr)

        if showcoord:
            buf.write(' (at %s)' % self.coord)
        buf.write('\n')

        for (child_name, child) in self.children():
            child.show(
                buf,
                offset=offset + 2,
                attrnames=attrnames,
                nodenames=nodenames,
                showcoord=showcoord,
                _my_node_name=child_name)


class NodeVisitor(object):
    """ A base NodeVisitor class for visiting c_ast nodes.
        Subclass it and define your own visit_XXX methods, where
        XXX is the class name you want to visit with these
        methods.

        For example:

        class ConstantVisitor(NodeVisitor):
            def __init__(self):
                self.values = []

            def visit_Constant(self, node):
                self.values.append(node.value)

        Creates a list of values of all the constant nodes
        encountered below the given node. To use it:

        cv = ConstantVisitor()
        cv.visit(node)

        Notes:

        *   generic_visit() will be called for AST nodes for which
            no visit_XXX method was defined.
        *   The children of nodes for which a visit_XXX was
            defined will not be visited - if you need this, call
            generic_visit() on the node.
            You can use:
                NodeVisitor.generic_visit(self, node)
        *   Modeled after Python's own AST visiting facilities
            (the ast module of Python 3.0)
    """

    _method_cache = None

    def visit(self, node):
        """ Visit a node.
        """

        if self._method_cache is None:
            self._method_cache = {}

        visitor = self._method_cache.get(node.__class__.__name__, None)
        if visitor is None:
            method = 'visit_' + node.__class__.__name__
            visitor = getattr(self, method, self.generic_visit)
            self._method_cache[node.__class__.__name__] = visitor

        return visitor(node)

    def generic_visit(self, node):
        """ Called if no explicit visitor function exists for a
            node. Implements preorder visiting of the node.
        """
        for c in node:
            self.visit(c)

'''

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\plotting\pygletplot\color_scheme.py ===
from sympy.core.basic import Basic
from sympy.core.symbol import (Symbol, symbols)
from sympy.utilities.lambdify import lambdify
from .util import interpolate, rinterpolate, create_bounds, update_bounds
from sympy.utilities.iterables import sift


class ColorGradient:
    colors = [0.4, 0.4, 0.4], [0.9, 0.9, 0.9]
    intervals = 0.0, 1.0

    def __init__(self, *args):
        if len(args) == 2:
            self.colors = list(args)
            self.intervals = [0.0, 1.0]
        elif len(args) > 0:
            if len(args) % 2 != 0:
                raise ValueError("len(args) should be even")
            self.colors = [args[i] for i in range(1, len(args), 2)]
            self.intervals = [args[i] for i in range(0, len(args), 2)]
        assert len(self.colors) == len(self.intervals)

    def copy(self):
        c = ColorGradient()
        c.colors = [e[::] for e in self.colors]
        c.intervals = self.intervals[::]
        return c

    def _find_interval(self, v):
        m = len(self.intervals)
        i = 0
        while i < m - 1 and self.intervals[i] <= v:
            i += 1
        return i

    def _interpolate_axis(self, axis, v):
        i = self._find_interval(v)
        v = rinterpolate(self.intervals[i - 1], self.intervals[i], v)
        return interpolate(self.colors[i - 1][axis], self.colors[i][axis], v)

    def __call__(self, r, g, b):
        c = self._interpolate_axis
        return c(0, r), c(1, g), c(2, b)

default_color_schemes = {}  # defined at the bottom of this file


class ColorScheme:

    def __init__(self, *args, **kwargs):
        self.args = args
        self.f, self.gradient = None, ColorGradient()

        if len(args) == 1 and not isinstance(args[0], Basic) and callable(args[0]):
            self.f = args[0]
        elif len(args) == 1 and isinstance(args[0], str):
            if args[0] in default_color_schemes:
                cs = default_color_schemes[args[0]]
                self.f, self.gradient = cs.f, cs.gradient.copy()
            else:
                self.f = lambdify('x,y,z,u,v', args[0])
        else:
            self.f, self.gradient = self._interpret_args(args)
        self._test_color_function()
        if not isinstance(self.gradient, ColorGradient):
            raise ValueError("Color gradient not properly initialized. "
                             "(Not a ColorGradient instance.)")

    def _interpret_args(self, args):
        f, gradient = None, self.gradient
        atoms, lists = self._sort_args(args)
        s = self._pop_symbol_list(lists)
        s = self._fill_in_vars(s)

        # prepare the error message for lambdification failure
        f_str = ', '.join(str(fa) for fa in atoms)
        s_str = (str(sa) for sa in s)
        s_str = ', '.join(sa for sa in s_str if sa.find('unbound') < 0)
        f_error = ValueError("Could not interpret arguments "
                             "%s as functions of %s." % (f_str, s_str))

        # try to lambdify args
        if len(atoms) == 1:
            fv = atoms[0]
            try:
                f = lambdify(s, [fv, fv, fv])
            except TypeError:
                raise f_error

        elif len(atoms) == 3:
            fr, fg, fb = atoms
            try:
                f = lambdify(s, [fr, fg, fb])
            except TypeError:
                raise f_error

        else:
            raise ValueError("A ColorScheme must provide 1 or 3 "
                             "functions in x, y, z, u, and/or v.")

        # try to intrepret any given color information
        if len(lists) == 0:
            gargs = []

        elif len(lists) == 1:
            gargs = lists[0]

        elif len(lists) == 2:
            try:
                (r1, g1, b1), (r2, g2, b2) = lists
            except TypeError:
                raise ValueError("If two color arguments are given, "
                                 "they must be given in the format "
                                 "(r1, g1, b1), (r2, g2, b2).")
            gargs = lists

        elif len(lists) == 3:
            try:
                (r1, r2), (g1, g2), (b1, b2) = lists
            except Exception:
                raise ValueError("If three color arguments are given, "
                                 "they must be given in the format "
                                 "(r1, r2), (g1, g2), (b1, b2). To create "
                                 "a multi-step gradient, use the syntax "
                                 "[0, colorStart, step1, color1, ..., 1, "
                                 "colorEnd].")
            gargs = [[r1, g1, b1], [r2, g2, b2]]

        else:
            raise ValueError("Don't know what to do with collection "
                             "arguments %s." % (', '.join(str(l) for l in lists)))

        if gargs:
            try:
                gradient = ColorGradient(*gargs)
            except Exception as ex:
                raise ValueError(("Could not initialize a gradient "
                                  "with arguments %s. Inner "
                                  "exception: %s") % (gargs, str(ex)))

        return f, gradient

    def _pop_symbol_list(self, lists):
        symbol_lists = []
        for l in lists:
            mark = True
            for s in l:
                if s is not None and not isinstance(s, Symbol):
                    mark = False
                    break
            if mark:
                lists.remove(l)
                symbol_lists.append(l)
        if len(symbol_lists) == 1:
            return symbol_lists[0]
        elif len(symbol_lists) == 0:
            return []
        else:
            raise ValueError("Only one list of Symbols "
                             "can be given for a color scheme.")

    def _fill_in_vars(self, args):
        defaults = symbols('x,y,z,u,v')
        v_error = ValueError("Could not find what to plot.")
        if len(args) == 0:
            return defaults
        if not isinstance(args, (tuple, list)):
            raise v_error
        if len(args) == 0:
            return defaults
        for s in args:
            if s is not None and not isinstance(s, Symbol):
                raise v_error
        # when vars are given explicitly, any vars
        # not given are marked 'unbound' as to not
        # be accidentally used in an expression
        vars = [Symbol('unbound%i' % (i)) for i in range(1, 6)]
        # interpret as t
        if len(args) == 1:
            vars[3] = args[0]
        # interpret as u,v
        elif len(args) == 2:
            if args[0] is not None:
                vars[3] = args[0]
            if args[1] is not None:
                vars[4] = args[1]
        # interpret as x,y,z
        elif len(args) >= 3:
            # allow some of x,y,z to be
            # left unbound if not given
            if args[0] is not None:
                vars[0] = args[0]
            if args[1] is not None:
                vars[1] = args[1]
            if args[2] is not None:
                vars[2] = args[2]
            # interpret the rest as t
            if len(args) >= 4:
                vars[3] = args[3]
                # ...or u,v
                if len(args) >= 5:
                    vars[4] = args[4]
        return vars

    def _sort_args(self, args):
        lists, atoms = sift(args,
            lambda a: isinstance(a, (tuple, list)), binary=True)
        return atoms, lists

    def _test_color_function(self):
        if not callable(self.f):
            raise ValueError("Color function is not callable.")
        try:
            result = self.f(0, 0, 0, 0, 0)
            if len(result) != 3:
                raise ValueError("length should be equal to 3")
        except TypeError:
            raise ValueError("Color function needs to accept x,y,z,u,v, "
                             "as arguments even if it doesn't use all of them.")
        except AssertionError:
            raise ValueError("Color function needs to return 3-tuple r,g,b.")
        except Exception:
            pass  # color function probably not valid at 0,0,0,0,0

    def __call__(self, x, y, z, u, v):
        try:
            return self.f(x, y, z, u, v)
        except Exception:
            return None

    def apply_to_curve(self, verts, u_set, set_len=None, inc_pos=None):
        """
        Apply this color scheme to a
        set of vertices over a single
        independent variable u.
        """
        bounds = create_bounds()
        cverts = []
        if callable(set_len):
            set_len(len(u_set)*2)
        # calculate f() = r,g,b for each vert
        # and find the min and max for r,g,b
        for _u in range(len(u_set)):
            if verts[_u] is None:
                cverts.append(None)
            else:
                x, y, z = verts[_u]
                u, v = u_set[_u], None
                c = self(x, y, z, u, v)
                if c is not None:
                    c = list(c)
                    update_bounds(bounds, c)
                cverts.append(c)
            if callable(inc_pos):
                inc_pos()
        # scale and apply gradient
        for _u in range(len(u_set)):
            if cverts[_u] is not None:
                for _c in range(3):
                    # scale from [f_min, f_max] to [0,1]
                    cverts[_u][_c] = rinterpolate(bounds[_c][0], bounds[_c][1],
                                                  cverts[_u][_c])
                # apply gradient
                cverts[_u] = self.gradient(*cverts[_u])
            if callable(inc_pos):
                inc_pos()
        return cverts

    def apply_to_surface(self, verts, u_set, v_set, set_len=None, inc_pos=None):
        """
        Apply this color scheme to a
        set of vertices over two
        independent variables u and v.
        """
        bounds = create_bounds()
        cverts = []
        if callable(set_len):
            set_len(len(u_set)*len(v_set)*2)
        # calculate f() = r,g,b for each vert
        # and find the min and max for r,g,b
        for _u in range(len(u_set)):
            column = []
            for _v in range(len(v_set)):
                if verts[_u][_v] is None:
                    column.append(None)
                else:
                    x, y, z = verts[_u][_v]
                    u, v = u_set[_u], v_set[_v]
                    c = self(x, y, z, u, v)
                    if c is not None:
                        c = list(c)
                        update_bounds(bounds, c)
                    column.append(c)
                if callable(inc_pos):
                    inc_pos()
            cverts.append(column)
        # scale and apply gradient
        for _u in range(len(u_set)):
            for _v in range(len(v_set)):
                if cverts[_u][_v] is not None:
                    # scale from [f_min, f_max] to [0,1]
                    for _c in range(3):
                        cverts[_u][_v][_c] = rinterpolate(bounds[_c][0],
                                             bounds[_c][1], cverts[_u][_v][_c])
                    # apply gradient
                    cverts[_u][_v] = self.gradient(*cverts[_u][_v])
                if callable(inc_pos):
                    inc_pos()
        return cverts

    def str_base(self):
        return ", ".join(str(a) for a in self.args)

    def __repr__(self):
        return "%s" % (self.str_base())


x, y, z, t, u, v = symbols('x,y,z,t,u,v')

default_color_schemes['rainbow'] = ColorScheme(z, y, x)
default_color_schemes['zfade'] = ColorScheme(z, (0.4, 0.4, 0.97),
                                             (0.97, 0.4, 0.4), (None, None, z))
default_color_schemes['zfade3'] = ColorScheme(z, (None, None, z),
                                              [0.00, (0.2, 0.2, 1.0),
                                               0.35, (0.2, 0.8, 0.4),
                                               0.50, (0.3, 0.9, 0.3),
                                               0.65, (0.4, 0.8, 0.2),
                                               1.00, (1.0, 0.2, 0.2)])

default_color_schemes['zfade4'] = ColorScheme(z, (None, None, z),
                                              [0.0, (0.3, 0.3, 1.0),
                                               0.30, (0.3, 1.0, 0.3),
                                               0.55, (0.95, 1.0, 0.2),
                                               0.65, (1.0, 0.95, 0.2),
                                               0.85, (1.0, 0.7, 0.2),
                                               1.0, (1.0, 0.3, 0.2)])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\test_highlevel_socket.py ===
from __future__ import annotations

import errno
import socket as stdlib_socket
import sys
from typing import TYPE_CHECKING

import pytest

from .. import _core, socket as tsocket
from .._highlevel_socket import *
from ..testing import (
    assert_checkpoints,
    check_half_closeable_stream,
    wait_all_tasks_blocked,
)
from .test_socket import setsockopt_tests

if TYPE_CHECKING:
    from collections.abc import Sequence


async def test_SocketStream_basics() -> None:
    # stdlib socket bad (even if connected)
    stdlib_a, stdlib_b = stdlib_socket.socketpair()
    with stdlib_a, stdlib_b:
        with pytest.raises(TypeError):
            SocketStream(stdlib_a)  # type: ignore[arg-type]

    # DGRAM socket bad
    with tsocket.socket(type=tsocket.SOCK_DGRAM) as sock:
        with pytest.raises(
            ValueError,
            match=r"^SocketStream requires a SOCK_STREAM socket$",
        ):
            # TODO: does not raise an error?
            SocketStream(sock)

    a, b = tsocket.socketpair()
    with a, b:
        s = SocketStream(a)
        assert s.socket is a

    # Use a real, connected socket to test socket options, because
    # socketpair() might give us a unix socket that doesn't support any of
    # these options
    with tsocket.socket() as listen_sock:
        await listen_sock.bind(("127.0.0.1", 0))
        listen_sock.listen(1)
        with tsocket.socket() as client_sock:
            await client_sock.connect(listen_sock.getsockname())

            s = SocketStream(client_sock)

            # TCP_NODELAY enabled by default
            assert s.getsockopt(tsocket.IPPROTO_TCP, tsocket.TCP_NODELAY)
            # We can disable it though
            s.setsockopt(tsocket.IPPROTO_TCP, tsocket.TCP_NODELAY, False)
            assert not s.getsockopt(tsocket.IPPROTO_TCP, tsocket.TCP_NODELAY)

            res = s.getsockopt(tsocket.IPPROTO_TCP, tsocket.TCP_NODELAY, 1)
            assert isinstance(res, bytes)

            setsockopt_tests(s)


async def test_SocketStream_send_all() -> None:
    BIG = 10000000

    a_sock, b_sock = tsocket.socketpair()
    with a_sock, b_sock:
        a = SocketStream(a_sock)
        b = SocketStream(b_sock)

        # Check a send_all that has to be split into multiple parts (on most
        # platforms... on Windows every send() either succeeds or fails as a
        # whole)
        async def sender() -> None:
            data = bytearray(BIG)
            await a.send_all(data)
            # send_all uses memoryviews internally, which temporarily "lock"
            # the object they view. If it doesn't clean them up properly, then
            # some bytearray operations might raise an error afterwards, which
            # would be a pretty weird and annoying side-effect to spring on
            # users. So test that this doesn't happen, by forcing the
            # bytearray's underlying buffer to be realloc'ed:
            data += bytes(BIG)
            # (Note: the above line of code doesn't do a very good job at
            # testing anything, because:
            # - on CPython, the refcount GC generally cleans up memoryviews
            #   for us even if we're sloppy.
            # - on PyPy3, at least as of 5.7.0, the memoryview code and the
            #   bytearray code conspire so that resizing never fails – if
            #   resizing forces the bytearray's internal buffer to move, then
            #   all memoryview references are automagically updated (!!).
            #   See:
            #   https://gist.github.com/njsmith/0ffd38ec05ad8e34004f34a7dc492227
            # But I'm leaving the test here in hopes that if this ever changes
            # and we break our implementation of send_all, then we'll get some
            # early warning...)

        async def receiver() -> None:
            # Make sure the sender fills up the kernel buffers and blocks
            await wait_all_tasks_blocked()
            nbytes = 0
            while nbytes < BIG:
                nbytes += len(await b.receive_some(BIG))
            assert nbytes == BIG

        async with _core.open_nursery() as nursery:
            nursery.start_soon(sender)
            nursery.start_soon(receiver)

        # We know that we received BIG bytes of NULs so far. Make sure that
        # was all the data in there.
        await a.send_all(b"e")
        assert await b.receive_some(10) == b"e"
        await a.send_eof()
        assert await b.receive_some(10) == b""


async def fill_stream(s: SocketStream) -> None:
    async def sender() -> None:
        while True:
            await s.send_all(b"x" * 10000)

    async def waiter(nursery: _core.Nursery) -> None:
        await wait_all_tasks_blocked()
        nursery.cancel_scope.cancel()

    async with _core.open_nursery() as nursery:
        nursery.start_soon(sender)
        nursery.start_soon(waiter, nursery)


async def test_SocketStream_generic() -> None:
    async def stream_maker() -> tuple[
        SocketStream,
        SocketStream,
    ]:
        left, right = tsocket.socketpair()
        return SocketStream(left), SocketStream(right)

    async def clogged_stream_maker() -> tuple[SocketStream, SocketStream]:
        left, right = await stream_maker()
        await fill_stream(left)
        await fill_stream(right)
        return left, right

    await check_half_closeable_stream(stream_maker, clogged_stream_maker)


async def test_SocketListener() -> None:
    # Not a Trio socket
    with stdlib_socket.socket() as s:
        s.bind(("127.0.0.1", 0))
        s.listen(10)
        with pytest.raises(TypeError):
            SocketListener(s)  # type: ignore[arg-type]

    # Not a SOCK_STREAM
    with tsocket.socket(type=tsocket.SOCK_DGRAM) as s:
        await s.bind(("127.0.0.1", 0))
        with pytest.raises(
            ValueError,
            match=r"^SocketListener requires a SOCK_STREAM socket$",
        ) as excinfo:
            SocketListener(s)
        excinfo.match(r".*SOCK_STREAM")

    # Didn't call .listen()
    # macOS has no way to check for this, so skip testing it there.
    if sys.platform != "darwin":
        with tsocket.socket() as s:
            await s.bind(("127.0.0.1", 0))
            with pytest.raises(
                ValueError,
                match=r"^SocketListener requires a listening socket$",
            ) as excinfo:
                SocketListener(s)
            excinfo.match(r".*listen")

    listen_sock = tsocket.socket()
    await listen_sock.bind(("127.0.0.1", 0))
    listen_sock.listen(10)
    listener = SocketListener(listen_sock)

    assert listener.socket is listen_sock

    client_sock = tsocket.socket()
    await client_sock.connect(listen_sock.getsockname())
    with assert_checkpoints():
        server_stream = await listener.accept()
    assert isinstance(server_stream, SocketStream)
    assert server_stream.socket.getsockname() == listen_sock.getsockname()
    assert server_stream.socket.getpeername() == client_sock.getsockname()

    with assert_checkpoints():
        await listener.aclose()

    with assert_checkpoints():
        await listener.aclose()

    with assert_checkpoints():
        with pytest.raises(_core.ClosedResourceError):
            await listener.accept()

    client_sock.close()
    await server_stream.aclose()


async def test_SocketListener_socket_closed_underfoot() -> None:
    listen_sock = tsocket.socket()
    await listen_sock.bind(("127.0.0.1", 0))
    listen_sock.listen(10)
    listener = SocketListener(listen_sock)

    # Close the socket, not the listener
    listen_sock.close()

    # SocketListener gives correct error
    with assert_checkpoints():
        with pytest.raises(_core.ClosedResourceError):
            await listener.accept()


async def test_SocketListener_accept_errors() -> None:
    class FakeSocket(tsocket.SocketType):
        def __init__(self, events: Sequence[SocketType | BaseException]) -> None:
            self._events = iter(events)

        type = tsocket.SOCK_STREAM

        # Fool the check for SO_ACCEPTCONN in SocketListener.__init__
        @overload
        def getsockopt(self, /, level: int, optname: int) -> int: ...

        @overload
        def getsockopt(  # noqa: F811
            self,
            /,
            level: int,
            optname: int,
            buflen: int,
        ) -> bytes: ...

        def getsockopt(  # noqa: F811
            self,
            /,
            level: int,
            optname: int,
            buflen: int | None = None,
        ) -> int | bytes:
            return True

        @overload
        def setsockopt(
            self,
            /,
            level: int,
            optname: int,
            value: int | Buffer,
        ) -> None: ...

        @overload
        def setsockopt(  # noqa: F811
            self,
            /,
            level: int,
            optname: int,
            value: None,
            optlen: int,
        ) -> None: ...

        def setsockopt(  # noqa: F811
            self,
            /,
            level: int,
            optname: int,
            value: int | Buffer | None,
            optlen: int | None = None,
        ) -> None:
            pass

        async def accept(self) -> tuple[SocketType, object]:
            await _core.checkpoint()
            event = next(self._events)
            if isinstance(event, BaseException):
                raise event
            else:
                return event, None

    fake_server_sock = FakeSocket([])

    fake_listen_sock = FakeSocket(
        [
            OSError(errno.ECONNABORTED, "Connection aborted"),
            OSError(errno.EPERM, "Permission denied"),
            OSError(errno.EPROTO, "Bad protocol"),
            fake_server_sock,
            OSError(errno.EMFILE, "Out of file descriptors"),
            OSError(errno.EFAULT, "attempt to write to read-only memory"),
            OSError(errno.ENOBUFS, "out of buffers"),
            fake_server_sock,
        ],
    )

    listener = SocketListener(fake_listen_sock)

    with assert_checkpoints():
        stream = await listener.accept()
        assert stream.socket is fake_server_sock

    for code, match in {
        errno.EMFILE: r"\[\w+ \d+\] Out of file descriptors$",
        errno.EFAULT: r"\[\w+ \d+\] attempt to write to read-only memory$",
        errno.ENOBUFS: r"\[\w+ \d+\] out of buffers$",
    }.items():
        with assert_checkpoints():
            with pytest.raises(OSError, match=match) as excinfo:
                await listener.accept()
            assert excinfo.value.errno == code

    with assert_checkpoints():
        stream = await listener.accept()
        assert stream.socket is fake_server_sock


async def test_socket_stream_works_when_peer_has_already_closed() -> None:
    sock_a, sock_b = tsocket.socketpair()
    with sock_a, sock_b:
        await sock_b.send(b"x")
        sock_b.close()
        stream = SocketStream(sock_a)
        assert await stream.receive_some(1) == b"x"
        assert await stream.receive_some(1) == b""

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_attention_clip.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
from logging import getLogger

from fusion_attention import AttentionMask, FusionAttention
from fusion_options import AttentionMaskFormat
from onnx import NodeProto
from onnx_model import OnnxModel

logger = getLogger(__name__)


class FusionAttentionClip(FusionAttention):
    """
    Fuse Attention subgraph of Clip into one Attention node.
    """

    def __init__(
        self,
        model: OnnxModel,
        hidden_size: int,
        num_heads: int,
    ):
        attention_mask = AttentionMask(model)
        attention_mask.mask_format = AttentionMaskFormat.NoMask

        super().__init__(
            model,
            hidden_size,
            num_heads,
            attention_mask,
            use_multi_head_attention=False,
            search_op_types=["SkipLayerNormalization"],
        )

    def get_num_heads_and_hidden_size(self, reshape_q: NodeProto) -> tuple[int, int]:
        """Detect num_heads and hidden_size for ONNX model from MiDaS
        Args:
            reshape_q (NodeProto): reshape node for q
        Returns:
            Tuple[int, int]: num_heads and hidden_size
        """
        concat = self.model.match_parent(reshape_q, "Concat", 1)
        if concat is None or len(concat.input) != 4:
            return self.num_heads, self.hidden_size

        # The shape is a tensor like [?, ?, num_heads, head_size]
        num_head_value = self.model.get_constant_value(concat.input[2])
        if num_head_value is None:
            return self.num_heads, self.hidden_size  # Fall back to user specified value

        if len(num_head_value) != 1 or num_head_value[0] <= 0:
            return self.num_heads, self.hidden_size  # Fall back to user specified value

        num_heads = num_head_value[0]

        head_size_value = self.model.get_constant_value(concat.input[3])
        if head_size_value is None:
            return self.num_heads, self.hidden_size  # Fall back to user specified value

        if len(head_size_value) != 1 or head_size_value[0] <= 0:
            return self.num_heads, self.hidden_size  # Fall back to user specified value

        head_size = head_size_value[0]

        hidden_size = num_heads * head_size

        if self.num_heads > 0 and num_heads != self.num_heads:
            if self.num_heads_warning:
                logger.warning(f"--num_heads is {self.num_heads}. Detected value is {num_heads}. Using detected value.")
                self.num_heads_warning = False  # Do not show the warning more than once

        if self.hidden_size > 0 and hidden_size != self.hidden_size:
            if self.hidden_size_warning:
                logger.warning(
                    f"--hidden_size is {self.hidden_size}. Detected value is {hidden_size}. Using detected value."
                )
                self.hidden_size_warning = False  # Do not show the warning more than once

        return num_heads, hidden_size

    def fuse(self, normalize_node, input_name_to_nodes, output_name_to_node):
        skip_input_index = None
        node_before_layer_norm = None
        for i in [1, 0]:
            parent = self.model.match_parent(normalize_node, "SkipLayerNormalization", i)
            if parent is not None:
                skip_input_index = i
                node_before_layer_norm = parent

        root_input = None
        if node_before_layer_norm is not None:
            root_input = node_before_layer_norm.output[0]
        else:
            # Deal with the first attention after the embedding layer.
            for i in [0, 1]:
                node_before_layer_norm = None

                node_before_layer_norm_1 = self.model.match_parent(normalize_node, "Add", i)
                node_before_layer_norm_2 = self.model.match_parent(normalize_node, "LayerNormalization", i)
                if node_before_layer_norm_1 is not None:
                    #           Add -----------+
                    #            |             |
                    #        LayerNorm         |
                    #            |             |
                    #        LayerNorm         |
                    #            |             |
                    #   Attention subgraph     |
                    #            |             |
                    #      SkipLayerNorm ------+
                    node_before_layer_norm = node_before_layer_norm_1
                elif node_before_layer_norm_2 is not None:
                    #           Add
                    #            |
                    #        LayerNorm --------+
                    #            |             |
                    #        LayerNorm         |
                    #            |             |
                    #   Attention subgraph     |
                    #            |             |
                    #      SkipLayerNorm ------+
                    node_before_layer_norm = node_before_layer_norm_2

                if node_before_layer_norm is None:
                    continue
                child = self.model.find_first_child_by_type(
                    node_before_layer_norm,
                    "LayerNormalization",
                    input_name_to_nodes,
                    False,
                )
                if child is None:
                    continue
                root_input = child.output[0]
                skip_input_index = i
                break

            if skip_input_index is None:
                return

        qkv_nodes = self.model.match_parent_path(
            normalize_node,
            ["Add", "MatMul", "Reshape", "Transpose", "Reshape", "MatMul"],
            [1 - skip_input_index, None, None, 0, 0, 0],
        )
        if qkv_nodes is None:
            qkv_nodes = self.model.match_parent_path(
                normalize_node,
                ["Add", "MatMul", "Reshape", "Transpose", "MatMul"],
                [1, None, 0, 0, 0],
            )
            if qkv_nodes is None:
                logger.debug("fuse_attention: failed to match qkv path")
                return
        reshape_qkv, transpose_qkv, matmul_qkv = (
            qkv_nodes[2],
            qkv_nodes[3],
            qkv_nodes[-1],
        )

        v_nodes = self.model.match_parent_path(
            matmul_qkv,
            ["Reshape", "Transpose", "Reshape", "Add", "MatMul"],
            [1, 0, 0, 0, None],
        )
        if v_nodes is None:
            v_nodes = self.model.match_parent_path(
                matmul_qkv, ["Transpose", "Reshape", "Add", "MatMul"], [1, 0, 0, None]
            )
            if v_nodes is None:
                logger.debug("fuse_attention: failed to match v path")
                return

        add_v, matmul_v = v_nodes[-2], v_nodes[-1]

        causal_mask_input_index = None
        add_mask = None
        add_mask_indices = []
        qk_nodes = self.model.match_parent_path(
            matmul_qkv,
            ["Softmax", "Reshape", "Add", "Reshape", "MatMul"],
            [0, 0, 0, None, 0],
            return_indice=add_mask_indices,
        )
        if qk_nodes is None:
            qk_nodes = self.model.match_parent_path(
                matmul_qkv,
                ["Softmax", "MatMul"],
                [0, 0],
            )
            if qk_nodes is None:
                qk_nodes = self.model.match_parent_path(matmul_qkv, ["Softmax", "Add", "Mul", "MatMul"], [0, 0, 0, 0])
                if qk_nodes is not None:
                    add_mask = qk_nodes[1]
                else:
                    # If attention mask is not used, we can still match the qk path.
                    qk_nodes = self.model.match_parent_path(matmul_qkv, ["Softmax", "Mul", "MatMul"], [0, 0, 0])
                    if qk_nodes is None:
                        # Cast nodes are added in the model for fp16.
                        qk_nodes = self.model.match_parent_path(
                            matmul_qkv,
                            ["Cast", "Cast", "Softmax", "Add", "Mul", "MatMul"],
                            [0, 0, 0, 0, 0, 0],
                        )
                        if qk_nodes is not None:
                            add_mask = qk_nodes[3]
                        else:
                            # If attention mask is not used, we can still match the qk path.
                            qk_nodes = self.model.match_parent_path(
                                matmul_qkv,
                                ["Cast", "Cast", "Softmax", "Mul", "MatMul"],
                                [0, 0, 0, 0, 0],
                            )
                            if qk_nodes is None:
                                logger.debug("fuse_attention: failed to match qk path")
                                return
        else:
            assert len(add_mask_indices) == 1
            causal_mask_input_index = 1 - add_mask_indices[0]
            add_mask = qk_nodes[2]

        matmul_qk = qk_nodes[-1]

        q_nodes = self.model.match_parent_path(
            matmul_qk,
            ["Reshape", "Transpose", "Reshape", "Mul", "Add", "MatMul"],
            [0, 0, 0, 0, None, None],
        )
        if q_nodes is None:
            q_nodes = self.model.match_parent_path(
                matmul_qk, ["Transpose", "Reshape", "Add", "MatMul"], [0, 0, 0, None]
            )
            if q_nodes is None:
                logger.debug("fuse_attention: failed to match q path")
                return

            reshape_q = q_nodes[1]
        else:
            reshape_q = q_nodes[2]

        add_q, matmul_q = q_nodes[-2], q_nodes[-1]

        k_nodes = self.model.match_parent_path(
            matmul_qk,
            ["Transpose", "Reshape", "Transpose", "Reshape", "Add", "MatMul"],
            [1, 0, 0, 0, 0, None],
        )
        if k_nodes is None:
            k_nodes = self.model.match_parent_path(
                matmul_qk, ["Transpose", "Reshape", "Add", "MatMul"], [1, 0, 0, None]
            )
            if k_nodes is None:
                logger.debug("fuse_attention: failed to match k path")
                return

        add_k, matmul_k = k_nodes[-2], k_nodes[-1]

        if matmul_q.input[0] != root_input or matmul_k.input[0] != root_input or matmul_v.input[0] != root_input:
            logger.debug("fuse_attention: expect to have same input to q, k and v matmul")
            return

        num_heads, hidden_size = self.get_num_heads_and_hidden_size(reshape_q)
        if num_heads <= 0 or hidden_size <= 0:
            logger.debug("fuse_attention: failed to detect num_heads or hidden_size")
            return

        attention_last_node = reshape_qkv

        add_qk = ""
        if add_mask is not None:
            # 4D Add after Q x K'
            add_qk_nodes = self.model.match_parent_path(
                add_mask,
                [
                    "Where",
                    "Sub",
                    "Cast",
                    "Expand",
                    "Unsqueeze",
                    "Unsqueeze",
                    "Reshape",
                    "Reshape",
                    "Cast",
                ],
                [1, 2, 1, 0, 0, 0, 0, 0, 0],
            )
            if add_qk_nodes is not None:
                add_qk = add_mask.input[1]
            else:
                # Here we do not match the whole subgraph since it is very complex. Instead, we just check whether a key path
                # of computing causal mask.
                causal_mask_nodes_1 = self.model.match_parent_path(
                    add_mask,
                    ["Concat", "Expand", "Unsqueeze", "Unsqueeze", "Where", "Less"],
                    [causal_mask_input_index, 0, 0, 0, 0, 0],
                )
                # If the model is exported with batch_size == 1, there is no Concat node
                causal_mask_nodes_2 = self.model.match_parent_path(
                    add_mask,
                    ["Expand", "Unsqueeze", "Unsqueeze", "Where", "Less"],
                    [causal_mask_input_index, 0, 0, 0, 0],
                )
                if causal_mask_nodes_1 is None and causal_mask_nodes_2 is None:
                    logger.debug("fuse_attention: failed to match causal mask subgraph")
                    return

        new_node = self.create_attention_node(
            mask_index=None,
            q_matmul=matmul_q,
            k_matmul=matmul_k,
            v_matmul=matmul_v,
            q_add=add_q,
            k_add=add_k,
            v_add=add_v,
            num_heads=num_heads,
            hidden_size=hidden_size,
            first_input=root_input,
            output=attention_last_node.output[0],
            add_qk_str=add_qk,
            scale=None,
            causal=(add_mask is not None),
        )
        if new_node is None:
            logger.debug("fuse_attention: failed to create fused node")
            return

        self.nodes_to_add.append(new_node)
        self.node_name_to_graph_name[new_node.name] = self.this_graph_name
        self.nodes_to_remove.extend([attention_last_node, transpose_qkv])
        self.increase_counter(new_node.op_type)

        # Use prune graph to remove nodes since they are shared by all attention nodes.
        self.prune_graph = True

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\utils\unpacking.py ===
"""Utilities related archives."""

import logging
import os
import shutil
import stat
import sys
import tarfile
import zipfile
from typing import Iterable, List, Optional
from zipfile import ZipInfo

from pip._internal.exceptions import InstallationError
from pip._internal.utils.filetypes import (
    BZ2_EXTENSIONS,
    TAR_EXTENSIONS,
    XZ_EXTENSIONS,
    ZIP_EXTENSIONS,
)
from pip._internal.utils.misc import ensure_dir

logger = logging.getLogger(__name__)


SUPPORTED_EXTENSIONS = ZIP_EXTENSIONS + TAR_EXTENSIONS

try:
    import bz2  # noqa

    SUPPORTED_EXTENSIONS += BZ2_EXTENSIONS
except ImportError:
    logger.debug("bz2 module is not available")

try:
    # Only for Python 3.3+
    import lzma  # noqa

    SUPPORTED_EXTENSIONS += XZ_EXTENSIONS
except ImportError:
    logger.debug("lzma module is not available")


def current_umask() -> int:
    """Get the current umask which involves having to set it temporarily."""
    mask = os.umask(0)
    os.umask(mask)
    return mask


def split_leading_dir(path: str) -> List[str]:
    path = path.lstrip("/").lstrip("\\")
    if "/" in path and (
        ("\\" in path and path.find("/") < path.find("\\")) or "\\" not in path
    ):
        return path.split("/", 1)
    elif "\\" in path:
        return path.split("\\", 1)
    else:
        return [path, ""]


def has_leading_dir(paths: Iterable[str]) -> bool:
    """Returns true if all the paths have the same leading path name
    (i.e., everything is in one subdirectory in an archive)"""
    common_prefix = None
    for path in paths:
        prefix, rest = split_leading_dir(path)
        if not prefix:
            return False
        elif common_prefix is None:
            common_prefix = prefix
        elif prefix != common_prefix:
            return False
    return True


def is_within_directory(directory: str, target: str) -> bool:
    """
    Return true if the absolute path of target is within the directory
    """
    abs_directory = os.path.abspath(directory)
    abs_target = os.path.abspath(target)

    prefix = os.path.commonprefix([abs_directory, abs_target])
    return prefix == abs_directory


def _get_default_mode_plus_executable() -> int:
    return 0o777 & ~current_umask() | 0o111


def set_extracted_file_to_default_mode_plus_executable(path: str) -> None:
    """
    Make file present at path have execute for user/group/world
    (chmod +x) is no-op on windows per python docs
    """
    os.chmod(path, _get_default_mode_plus_executable())


def zip_item_is_executable(info: ZipInfo) -> bool:
    mode = info.external_attr >> 16
    # if mode and regular file and any execute permissions for
    # user/group/world?
    return bool(mode and stat.S_ISREG(mode) and mode & 0o111)


def unzip_file(filename: str, location: str, flatten: bool = True) -> None:
    """
    Unzip the file (with path `filename`) to the destination `location`.  All
    files are written based on system defaults and umask (i.e. permissions are
    not preserved), except that regular file members with any execute
    permissions (user, group, or world) have "chmod +x" applied after being
    written. Note that for windows, any execute changes using os.chmod are
    no-ops per the python docs.
    """
    ensure_dir(location)
    zipfp = open(filename, "rb")
    try:
        zip = zipfile.ZipFile(zipfp, allowZip64=True)
        leading = has_leading_dir(zip.namelist()) and flatten
        for info in zip.infolist():
            name = info.filename
            fn = name
            if leading:
                fn = split_leading_dir(name)[1]
            fn = os.path.join(location, fn)
            dir = os.path.dirname(fn)
            if not is_within_directory(location, fn):
                message = (
                    "The zip file ({}) has a file ({}) trying to install "
                    "outside target directory ({})"
                )
                raise InstallationError(message.format(filename, fn, location))
            if fn.endswith("/") or fn.endswith("\\"):
                # A directory
                ensure_dir(fn)
            else:
                ensure_dir(dir)
                # Don't use read() to avoid allocating an arbitrarily large
                # chunk of memory for the file's content
                fp = zip.open(name)
                try:
                    with open(fn, "wb") as destfp:
                        shutil.copyfileobj(fp, destfp)
                finally:
                    fp.close()
                    if zip_item_is_executable(info):
                        set_extracted_file_to_default_mode_plus_executable(fn)
    finally:
        zipfp.close()


def untar_file(filename: str, location: str) -> None:
    """
    Untar the file (with path `filename`) to the destination `location`.
    All files are written based on system defaults and umask (i.e. permissions
    are not preserved), except that regular file members with any execute
    permissions (user, group, or world) have "chmod +x" applied on top of the
    default.  Note that for windows, any execute changes using os.chmod are
    no-ops per the python docs.
    """
    ensure_dir(location)
    if filename.lower().endswith(".gz") or filename.lower().endswith(".tgz"):
        mode = "r:gz"
    elif filename.lower().endswith(BZ2_EXTENSIONS):
        mode = "r:bz2"
    elif filename.lower().endswith(XZ_EXTENSIONS):
        mode = "r:xz"
    elif filename.lower().endswith(".tar"):
        mode = "r"
    else:
        logger.warning(
            "Cannot determine compression type for file %s",
            filename,
        )
        mode = "r:*"

    tar = tarfile.open(filename, mode, encoding="utf-8")  # type: ignore
    try:
        leading = has_leading_dir([member.name for member in tar.getmembers()])

        # PEP 706 added `tarfile.data_filter`, and made some other changes to
        # Python's tarfile module (see below). The features were backported to
        # security releases.
        try:
            data_filter = tarfile.data_filter
        except AttributeError:
            _untar_without_filter(filename, location, tar, leading)
        else:
            default_mode_plus_executable = _get_default_mode_plus_executable()

            if leading:
                # Strip the leading directory from all files in the archive,
                # including hardlink targets (which are relative to the
                # unpack location).
                for member in tar.getmembers():
                    name_lead, name_rest = split_leading_dir(member.name)
                    member.name = name_rest
                    if member.islnk():
                        lnk_lead, lnk_rest = split_leading_dir(member.linkname)
                        if lnk_lead == name_lead:
                            member.linkname = lnk_rest

            def pip_filter(member: tarfile.TarInfo, path: str) -> tarfile.TarInfo:
                orig_mode = member.mode
                try:
                    try:
                        member = data_filter(member, location)
                    except tarfile.LinkOutsideDestinationError:
                        if sys.version_info[:3] in {
                            (3, 9, 17),
                            (3, 10, 12),
                            (3, 11, 4),
                        }:
                            # The tarfile filter in specific Python versions
                            # raises LinkOutsideDestinationError on valid input
                            # (https://github.com/python/cpython/issues/107845)
                            # Ignore the error there, but do use the
                            # more lax `tar_filter`
                            member = tarfile.tar_filter(member, location)
                        else:
                            raise
                except tarfile.TarError as exc:
                    message = "Invalid member in the tar file {}: {}"
                    # Filter error messages mention the member name.
                    # No need to add it here.
                    raise InstallationError(
                        message.format(
                            filename,
                            exc,
                        )
                    )
                if member.isfile() and orig_mode & 0o111:
                    member.mode = default_mode_plus_executable
                else:
                    # See PEP 706 note above.
                    # The PEP changed this from `int` to `Optional[int]`,
                    # where None means "use the default". Mypy doesn't
                    # know this yet.
                    member.mode = None  # type: ignore [assignment]
                return member

            tar.extractall(location, filter=pip_filter)

    finally:
        tar.close()


def _untar_without_filter(
    filename: str,
    location: str,
    tar: tarfile.TarFile,
    leading: bool,
) -> None:
    """Fallback for Python without tarfile.data_filter"""
    for member in tar.getmembers():
        fn = member.name
        if leading:
            fn = split_leading_dir(fn)[1]
        path = os.path.join(location, fn)
        if not is_within_directory(location, path):
            message = (
                "The tar file ({}) has a file ({}) trying to install "
                "outside target directory ({})"
            )
            raise InstallationError(message.format(filename, path, location))
        if member.isdir():
            ensure_dir(path)
        elif member.issym():
            try:
                tar._extract_member(member, path)
            except Exception as exc:
                # Some corrupt tar files seem to produce this
                # (specifically bad symlinks)
                logger.warning(
                    "In the tar file %s the member %s is invalid: %s",
                    filename,
                    member.name,
                    exc,
                )
                continue
        else:
            try:
                fp = tar.extractfile(member)
            except (KeyError, AttributeError) as exc:
                # Some corrupt tar files seem to produce this
                # (specifically bad symlinks)
                logger.warning(
                    "In the tar file %s the member %s is invalid: %s",
                    filename,
                    member.name,
                    exc,
                )
                continue
            ensure_dir(os.path.dirname(path))
            assert fp is not None
            with open(path, "wb") as destfp:
                shutil.copyfileobj(fp, destfp)
            fp.close()
            # Update the timestamp (useful for cython compiled files)
            tar.utime(member, path)
            # member have any execute permissions for user/group/world?
            if member.mode & 0o111:
                set_extracted_file_to_default_mode_plus_executable(path)


def unpack_file(
    filename: str,
    location: str,
    content_type: Optional[str] = None,
) -> None:
    filename = os.path.realpath(filename)
    if (
        content_type == "application/zip"
        or filename.lower().endswith(ZIP_EXTENSIONS)
        or zipfile.is_zipfile(filename)
    ):
        unzip_file(filename, location, flatten=not filename.endswith(".whl"))
    elif (
        content_type == "application/x-gzip"
        or tarfile.is_tarfile(filename)
        or filename.lower().endswith(TAR_EXTENSIONS + BZ2_EXTENSIONS + XZ_EXTENSIONS)
    ):
        untar_file(filename, location)
    else:
        # FIXME: handle?
        # FIXME: magic signatures?
        logger.critical(
            "Cannot unpack file %s (downloaded from %s, content-type: %s); "
            "cannot detect archive format",
            filename,
            location,
            content_type,
        )
        raise InstallationError(f"Cannot determine archive format of {location}")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\mysql\aiomysql.py ===
# dialects/mysql/aiomysql.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors <see AUTHORS
# file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors

r"""
.. dialect:: mysql+aiomysql
    :name: aiomysql
    :dbapi: aiomysql
    :connectstring: mysql+aiomysql://user:password@host:port/dbname[?key=value&key=value...]
    :url: https://github.com/aio-libs/aiomysql

The aiomysql dialect is SQLAlchemy's second Python asyncio dialect.

Using a special asyncio mediation layer, the aiomysql dialect is usable
as the backend for the :ref:`SQLAlchemy asyncio <asyncio_toplevel>`
extension package.

This dialect should normally be used only with the
:func:`_asyncio.create_async_engine` engine creation function::

    from sqlalchemy.ext.asyncio import create_async_engine

    engine = create_async_engine(
        "mysql+aiomysql://user:pass@hostname/dbname?charset=utf8mb4"
    )

"""  # noqa
from collections import deque

from .pymysql import MySQLDialect_pymysql
from ... import pool
from ... import util
from ...engine import AdaptedConnection
from ...util.concurrency import asyncio
from ...util.concurrency import await_fallback
from ...util.concurrency import await_only


class AsyncAdapt_aiomysql_cursor:
    # TODO: base on connectors/asyncio.py
    # see #10415
    server_side = False
    __slots__ = (
        "_adapt_connection",
        "_connection",
        "await_",
        "_cursor",
        "_rows",
    )

    def __init__(self, adapt_connection):
        self._adapt_connection = adapt_connection
        self._connection = adapt_connection._connection
        self.await_ = adapt_connection.await_

        cursor = self._connection.cursor(adapt_connection.dbapi.Cursor)

        # see https://github.com/aio-libs/aiomysql/issues/543
        self._cursor = self.await_(cursor.__aenter__())
        self._rows = deque()

    @property
    def description(self):
        return self._cursor.description

    @property
    def rowcount(self):
        return self._cursor.rowcount

    @property
    def arraysize(self):
        return self._cursor.arraysize

    @arraysize.setter
    def arraysize(self, value):
        self._cursor.arraysize = value

    @property
    def lastrowid(self):
        return self._cursor.lastrowid

    def close(self):
        # note we aren't actually closing the cursor here,
        # we are just letting GC do it.   to allow this to be async
        # we would need the Result to change how it does "Safe close cursor".
        # MySQL "cursors" don't actually have state to be "closed" besides
        # exhausting rows, which we already have done for sync cursor.
        # another option would be to emulate aiosqlite dialect and assign
        # cursor only if we are doing server side cursor operation.
        self._rows.clear()

    def execute(self, operation, parameters=None):
        return self.await_(self._execute_async(operation, parameters))

    def executemany(self, operation, seq_of_parameters):
        return self.await_(
            self._executemany_async(operation, seq_of_parameters)
        )

    async def _execute_async(self, operation, parameters):
        async with self._adapt_connection._execute_mutex:
            result = await self._cursor.execute(operation, parameters)

            if not self.server_side:
                # aiomysql has a "fake" async result, so we have to pull it out
                # of that here since our default result is not async.
                # we could just as easily grab "_rows" here and be done with it
                # but this is safer.
                self._rows = deque(await self._cursor.fetchall())
            return result

    async def _executemany_async(self, operation, seq_of_parameters):
        async with self._adapt_connection._execute_mutex:
            return await self._cursor.executemany(operation, seq_of_parameters)

    def setinputsizes(self, *inputsizes):
        pass

    def __iter__(self):
        while self._rows:
            yield self._rows.popleft()

    def fetchone(self):
        if self._rows:
            return self._rows.popleft()
        else:
            return None

    def fetchmany(self, size=None):
        if size is None:
            size = self.arraysize

        rr = self._rows
        return [rr.popleft() for _ in range(min(size, len(rr)))]

    def fetchall(self):
        retval = list(self._rows)
        self._rows.clear()
        return retval


class AsyncAdapt_aiomysql_ss_cursor(AsyncAdapt_aiomysql_cursor):
    # TODO: base on connectors/asyncio.py
    # see #10415
    __slots__ = ()
    server_side = True

    def __init__(self, adapt_connection):
        self._adapt_connection = adapt_connection
        self._connection = adapt_connection._connection
        self.await_ = adapt_connection.await_

        cursor = self._connection.cursor(adapt_connection.dbapi.SSCursor)

        self._cursor = self.await_(cursor.__aenter__())

    def close(self):
        if self._cursor is not None:
            self.await_(self._cursor.close())
            self._cursor = None

    def fetchone(self):
        return self.await_(self._cursor.fetchone())

    def fetchmany(self, size=None):
        return self.await_(self._cursor.fetchmany(size=size))

    def fetchall(self):
        return self.await_(self._cursor.fetchall())


class AsyncAdapt_aiomysql_connection(AdaptedConnection):
    # TODO: base on connectors/asyncio.py
    # see #10415
    await_ = staticmethod(await_only)
    __slots__ = ("dbapi", "_execute_mutex")

    def __init__(self, dbapi, connection):
        self.dbapi = dbapi
        self._connection = connection
        self._execute_mutex = asyncio.Lock()

    def ping(self, reconnect):
        return self.await_(self._connection.ping(reconnect))

    def character_set_name(self):
        return self._connection.character_set_name()

    def autocommit(self, value):
        self.await_(self._connection.autocommit(value))

    def cursor(self, server_side=False):
        if server_side:
            return AsyncAdapt_aiomysql_ss_cursor(self)
        else:
            return AsyncAdapt_aiomysql_cursor(self)

    def rollback(self):
        self.await_(self._connection.rollback())

    def commit(self):
        self.await_(self._connection.commit())

    def terminate(self):
        # it's not awaitable.
        self._connection.close()

    def close(self) -> None:
        self.await_(self._connection.ensure_closed())


class AsyncAdaptFallback_aiomysql_connection(AsyncAdapt_aiomysql_connection):
    # TODO: base on connectors/asyncio.py
    # see #10415
    __slots__ = ()

    await_ = staticmethod(await_fallback)


class AsyncAdapt_aiomysql_dbapi:
    def __init__(self, aiomysql, pymysql):
        self.aiomysql = aiomysql
        self.pymysql = pymysql
        self.paramstyle = "format"
        self._init_dbapi_attributes()
        self.Cursor, self.SSCursor = self._init_cursors_subclasses()

    def _init_dbapi_attributes(self):
        for name in (
            "Warning",
            "Error",
            "InterfaceError",
            "DataError",
            "DatabaseError",
            "OperationalError",
            "InterfaceError",
            "IntegrityError",
            "ProgrammingError",
            "InternalError",
            "NotSupportedError",
        ):
            setattr(self, name, getattr(self.aiomysql, name))

        for name in (
            "NUMBER",
            "STRING",
            "DATETIME",
            "BINARY",
            "TIMESTAMP",
            "Binary",
        ):
            setattr(self, name, getattr(self.pymysql, name))

    def connect(self, *arg, **kw):
        async_fallback = kw.pop("async_fallback", False)
        creator_fn = kw.pop("async_creator_fn", self.aiomysql.connect)

        if util.asbool(async_fallback):
            return AsyncAdaptFallback_aiomysql_connection(
                self,
                await_fallback(creator_fn(*arg, **kw)),
            )
        else:
            return AsyncAdapt_aiomysql_connection(
                self,
                await_only(creator_fn(*arg, **kw)),
            )

    def _init_cursors_subclasses(self):
        # suppress unconditional warning emitted by aiomysql
        class Cursor(self.aiomysql.Cursor):
            async def _show_warnings(self, conn):
                pass

        class SSCursor(self.aiomysql.SSCursor):
            async def _show_warnings(self, conn):
                pass

        return Cursor, SSCursor


class MySQLDialect_aiomysql(MySQLDialect_pymysql):
    driver = "aiomysql"
    supports_statement_cache = True

    supports_server_side_cursors = True
    _sscursor = AsyncAdapt_aiomysql_ss_cursor

    is_async = True
    has_terminate = True

    @classmethod
    def import_dbapi(cls):
        return AsyncAdapt_aiomysql_dbapi(
            __import__("aiomysql"), __import__("pymysql")
        )

    @classmethod
    def get_pool_class(cls, url):
        async_fallback = url.query.get("async_fallback", False)

        if util.asbool(async_fallback):
            return pool.FallbackAsyncAdaptedQueuePool
        else:
            return pool.AsyncAdaptedQueuePool

    def do_terminate(self, dbapi_connection) -> None:
        dbapi_connection.terminate()

    def create_connect_args(self, url):
        return super().create_connect_args(
            url, _translate_args=dict(username="user", database="db")
        )

    def is_disconnect(self, e, connection, cursor):
        if super().is_disconnect(e, connection, cursor):
            return True
        else:
            str_e = str(e).lower()
            return "not connected" in str_e

    def _found_rows_client_flag(self):
        from pymysql.constants import CLIENT

        return CLIENT.FOUND_ROWS

    def get_driver_connection(self, connection):
        return connection._connection


dialect = MySQLDialect_aiomysql

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\ext\mypy\names.py ===
# ext/mypy/names.py
# Copyright (C) 2021-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

from __future__ import annotations

from typing import Dict
from typing import List
from typing import Optional
from typing import Set
from typing import Tuple
from typing import Union

from mypy.nodes import ARG_POS
from mypy.nodes import CallExpr
from mypy.nodes import ClassDef
from mypy.nodes import Decorator
from mypy.nodes import Expression
from mypy.nodes import FuncDef
from mypy.nodes import MemberExpr
from mypy.nodes import NameExpr
from mypy.nodes import OverloadedFuncDef
from mypy.nodes import SymbolNode
from mypy.nodes import TypeAlias
from mypy.nodes import TypeInfo
from mypy.plugin import SemanticAnalyzerPluginInterface
from mypy.types import CallableType
from mypy.types import get_proper_type
from mypy.types import Instance
from mypy.types import UnboundType

from ... import util

COLUMN: int = util.symbol("COLUMN")
RELATIONSHIP: int = util.symbol("RELATIONSHIP")
REGISTRY: int = util.symbol("REGISTRY")
COLUMN_PROPERTY: int = util.symbol("COLUMN_PROPERTY")
TYPEENGINE: int = util.symbol("TYPEENGNE")
MAPPED: int = util.symbol("MAPPED")
DECLARATIVE_BASE: int = util.symbol("DECLARATIVE_BASE")
DECLARATIVE_META: int = util.symbol("DECLARATIVE_META")
MAPPED_DECORATOR: int = util.symbol("MAPPED_DECORATOR")
SYNONYM_PROPERTY: int = util.symbol("SYNONYM_PROPERTY")
COMPOSITE_PROPERTY: int = util.symbol("COMPOSITE_PROPERTY")
DECLARED_ATTR: int = util.symbol("DECLARED_ATTR")
MAPPER_PROPERTY: int = util.symbol("MAPPER_PROPERTY")
AS_DECLARATIVE: int = util.symbol("AS_DECLARATIVE")
AS_DECLARATIVE_BASE: int = util.symbol("AS_DECLARATIVE_BASE")
DECLARATIVE_MIXIN: int = util.symbol("DECLARATIVE_MIXIN")
QUERY_EXPRESSION: int = util.symbol("QUERY_EXPRESSION")

# names that must succeed with mypy.api.named_type
NAMED_TYPE_BUILTINS_OBJECT = "builtins.object"
NAMED_TYPE_BUILTINS_STR = "builtins.str"
NAMED_TYPE_BUILTINS_LIST = "builtins.list"
NAMED_TYPE_SQLA_MAPPED = "sqlalchemy.orm.base.Mapped"

_RelFullNames = {
    "sqlalchemy.orm.relationships.Relationship",
    "sqlalchemy.orm.relationships.RelationshipProperty",
    "sqlalchemy.orm.relationships._RelationshipDeclared",
    "sqlalchemy.orm.Relationship",
    "sqlalchemy.orm.RelationshipProperty",
}

_lookup: Dict[str, Tuple[int, Set[str]]] = {
    "Column": (
        COLUMN,
        {
            "sqlalchemy.sql.schema.Column",
            "sqlalchemy.sql.Column",
        },
    ),
    "Relationship": (RELATIONSHIP, _RelFullNames),
    "RelationshipProperty": (RELATIONSHIP, _RelFullNames),
    "_RelationshipDeclared": (RELATIONSHIP, _RelFullNames),
    "registry": (
        REGISTRY,
        {
            "sqlalchemy.orm.decl_api.registry",
            "sqlalchemy.orm.registry",
        },
    ),
    "ColumnProperty": (
        COLUMN_PROPERTY,
        {
            "sqlalchemy.orm.properties.MappedSQLExpression",
            "sqlalchemy.orm.MappedSQLExpression",
            "sqlalchemy.orm.properties.ColumnProperty",
            "sqlalchemy.orm.ColumnProperty",
        },
    ),
    "MappedSQLExpression": (
        COLUMN_PROPERTY,
        {
            "sqlalchemy.orm.properties.MappedSQLExpression",
            "sqlalchemy.orm.MappedSQLExpression",
            "sqlalchemy.orm.properties.ColumnProperty",
            "sqlalchemy.orm.ColumnProperty",
        },
    ),
    "Synonym": (
        SYNONYM_PROPERTY,
        {
            "sqlalchemy.orm.descriptor_props.Synonym",
            "sqlalchemy.orm.Synonym",
            "sqlalchemy.orm.descriptor_props.SynonymProperty",
            "sqlalchemy.orm.SynonymProperty",
        },
    ),
    "SynonymProperty": (
        SYNONYM_PROPERTY,
        {
            "sqlalchemy.orm.descriptor_props.Synonym",
            "sqlalchemy.orm.Synonym",
            "sqlalchemy.orm.descriptor_props.SynonymProperty",
            "sqlalchemy.orm.SynonymProperty",
        },
    ),
    "Composite": (
        COMPOSITE_PROPERTY,
        {
            "sqlalchemy.orm.descriptor_props.Composite",
            "sqlalchemy.orm.Composite",
            "sqlalchemy.orm.descriptor_props.CompositeProperty",
            "sqlalchemy.orm.CompositeProperty",
        },
    ),
    "CompositeProperty": (
        COMPOSITE_PROPERTY,
        {
            "sqlalchemy.orm.descriptor_props.Composite",
            "sqlalchemy.orm.Composite",
            "sqlalchemy.orm.descriptor_props.CompositeProperty",
            "sqlalchemy.orm.CompositeProperty",
        },
    ),
    "MapperProperty": (
        MAPPER_PROPERTY,
        {
            "sqlalchemy.orm.interfaces.MapperProperty",
            "sqlalchemy.orm.MapperProperty",
        },
    ),
    "TypeEngine": (TYPEENGINE, {"sqlalchemy.sql.type_api.TypeEngine"}),
    "Mapped": (MAPPED, {NAMED_TYPE_SQLA_MAPPED}),
    "declarative_base": (
        DECLARATIVE_BASE,
        {
            "sqlalchemy.ext.declarative.declarative_base",
            "sqlalchemy.orm.declarative_base",
            "sqlalchemy.orm.decl_api.declarative_base",
        },
    ),
    "DeclarativeMeta": (
        DECLARATIVE_META,
        {
            "sqlalchemy.ext.declarative.DeclarativeMeta",
            "sqlalchemy.orm.DeclarativeMeta",
            "sqlalchemy.orm.decl_api.DeclarativeMeta",
        },
    ),
    "mapped": (
        MAPPED_DECORATOR,
        {
            "sqlalchemy.orm.decl_api.registry.mapped",
            "sqlalchemy.orm.registry.mapped",
        },
    ),
    "as_declarative": (
        AS_DECLARATIVE,
        {
            "sqlalchemy.ext.declarative.as_declarative",
            "sqlalchemy.orm.decl_api.as_declarative",
            "sqlalchemy.orm.as_declarative",
        },
    ),
    "as_declarative_base": (
        AS_DECLARATIVE_BASE,
        {
            "sqlalchemy.orm.decl_api.registry.as_declarative_base",
            "sqlalchemy.orm.registry.as_declarative_base",
        },
    ),
    "declared_attr": (
        DECLARED_ATTR,
        {
            "sqlalchemy.orm.decl_api.declared_attr",
            "sqlalchemy.orm.declared_attr",
        },
    ),
    "declarative_mixin": (
        DECLARATIVE_MIXIN,
        {
            "sqlalchemy.orm.decl_api.declarative_mixin",
            "sqlalchemy.orm.declarative_mixin",
        },
    ),
    "query_expression": (
        QUERY_EXPRESSION,
        {
            "sqlalchemy.orm.query_expression",
            "sqlalchemy.orm._orm_constructors.query_expression",
        },
    ),
}


def has_base_type_id(info: TypeInfo, type_id: int) -> bool:
    for mr in info.mro:
        check_type_id, fullnames = _lookup.get(mr.name, (None, None))
        if check_type_id == type_id:
            break
    else:
        return False

    if fullnames is None:
        return False

    return mr.fullname in fullnames


def mro_has_id(mro: List[TypeInfo], type_id: int) -> bool:
    for mr in mro:
        check_type_id, fullnames = _lookup.get(mr.name, (None, None))
        if check_type_id == type_id:
            break
    else:
        return False

    if fullnames is None:
        return False

    return mr.fullname in fullnames


def type_id_for_unbound_type(
    type_: UnboundType, cls: ClassDef, api: SemanticAnalyzerPluginInterface
) -> Optional[int]:
    sym = api.lookup_qualified(type_.name, type_)
    if sym is not None:
        if isinstance(sym.node, TypeAlias):
            target_type = get_proper_type(sym.node.target)
            if isinstance(target_type, Instance):
                return type_id_for_named_node(target_type.type)
        elif isinstance(sym.node, TypeInfo):
            return type_id_for_named_node(sym.node)

    return None


def type_id_for_callee(callee: Expression) -> Optional[int]:
    if isinstance(callee, (MemberExpr, NameExpr)):
        if isinstance(callee.node, Decorator) and isinstance(
            callee.node.func, FuncDef
        ):
            if callee.node.func.type and isinstance(
                callee.node.func.type, CallableType
            ):
                ret_type = get_proper_type(callee.node.func.type.ret_type)

                if isinstance(ret_type, Instance):
                    return type_id_for_fullname(ret_type.type.fullname)

            return None

        elif isinstance(callee.node, OverloadedFuncDef):
            if (
                callee.node.impl
                and callee.node.impl.type
                and isinstance(callee.node.impl.type, CallableType)
            ):
                ret_type = get_proper_type(callee.node.impl.type.ret_type)

                if isinstance(ret_type, Instance):
                    return type_id_for_fullname(ret_type.type.fullname)

            return None
        elif isinstance(callee.node, FuncDef):
            if callee.node.type and isinstance(callee.node.type, CallableType):
                ret_type = get_proper_type(callee.node.type.ret_type)

                if isinstance(ret_type, Instance):
                    return type_id_for_fullname(ret_type.type.fullname)

            return None
        elif isinstance(callee.node, TypeAlias):
            target_type = get_proper_type(callee.node.target)
            if isinstance(target_type, Instance):
                return type_id_for_fullname(target_type.type.fullname)
        elif isinstance(callee.node, TypeInfo):
            return type_id_for_named_node(callee)
    return None


def type_id_for_named_node(
    node: Union[NameExpr, MemberExpr, SymbolNode]
) -> Optional[int]:
    type_id, fullnames = _lookup.get(node.name, (None, None))

    if type_id is None or fullnames is None:
        return None
    elif node.fullname in fullnames:
        return type_id
    else:
        return None


def type_id_for_fullname(fullname: str) -> Optional[int]:
    tokens = fullname.split(".")
    immediate = tokens[-1]

    type_id, fullnames = _lookup.get(immediate, (None, None))

    if type_id is None or fullnames is None:
        return None
    elif fullname in fullnames:
        return type_id
    else:
        return None


def expr_to_mapped_constructor(expr: Expression) -> CallExpr:
    column_descriptor = NameExpr("__sa_Mapped")
    column_descriptor.fullname = NAMED_TYPE_SQLA_MAPPED
    member_expr = MemberExpr(column_descriptor, "_empty_constructor")
    return CallExpr(
        member_expr,
        [expr],
        [ARG_POS],
        ["arg1"],
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\integrals\trigonometry.py ===
from sympy.core import cacheit, Dummy, Ne, Integer, Rational, S, Wild
from sympy.functions import binomial, sin, cos, Piecewise, Abs
from .integrals import integrate

# TODO sin(a*x)*cos(b*x) -> sin((a+b)x) + sin((a-b)x) ?

# creating, each time, Wild's and sin/cos/Mul is expensive. Also, our match &
# subs are very slow when not cached, and if we create Wild each time, we
# effectively block caching.
#
# so we cache the pattern

# need to use a function instead of lamda since hash of lambda changes on
# each call to _pat_sincos
def _integer_instance(n):
    return isinstance(n, Integer)

@cacheit
def _pat_sincos(x):
    a = Wild('a', exclude=[x])
    n, m = [Wild(s, exclude=[x], properties=[_integer_instance])
                for s in 'nm']
    pat = sin(a*x)**n * cos(a*x)**m
    return pat, a, n, m

_u = Dummy('u')


def trigintegrate(f, x, conds='piecewise'):
    """
    Integrate f = Mul(trig) over x.

    Examples
    ========

    >>> from sympy import sin, cos, tan, sec
    >>> from sympy.integrals.trigonometry import trigintegrate
    >>> from sympy.abc import x

    >>> trigintegrate(sin(x)*cos(x), x)
    sin(x)**2/2

    >>> trigintegrate(sin(x)**2, x)
    x/2 - sin(x)*cos(x)/2

    >>> trigintegrate(tan(x)*sec(x), x)
    1/cos(x)

    >>> trigintegrate(sin(x)*tan(x), x)
    -log(sin(x) - 1)/2 + log(sin(x) + 1)/2 - sin(x)

    References
    ==========

    .. [1] https://en.wikibooks.org/wiki/Calculus/Integration_techniques

    See Also
    ========

    sympy.integrals.integrals.Integral.doit
    sympy.integrals.integrals.Integral
    """
    pat, a, n, m = _pat_sincos(x)

    f = f.rewrite('sincos')
    M = f.match(pat)

    if M is None:
        return

    n, m = M[n], M[m]
    if n.is_zero and m.is_zero:
        return x
    zz = x if n.is_zero else S.Zero

    a = M[a]

    if n.is_odd or m.is_odd:
        u = _u
        n_, m_ = n.is_odd, m.is_odd

        # take smallest n or m -- to choose simplest substitution
        if n_ and m_:

            # Make sure to choose the positive one
            # otherwise an incorrect integral can occur.
            if n < 0 and m > 0:
                m_ = True
                n_ = False
            elif m < 0 and n > 0:
                n_ = True
                m_ = False
            # Both are negative so choose the smallest n or m
            # in absolute value for simplest substitution.
            elif (n < 0 and m < 0):
                n_ = n > m
                m_ = not (n > m)

            # Both n and m are odd and positive
            else:
                n_ = (n < m)      # NB: careful here, one of the
                m_ = not (n < m)  # conditions *must* be true

        #  n      m       u=C        (n-1)/2    m
        # S(x) * C(x) dx  --> -(1-u^2)       * u  du
        if n_:
            ff = -(1 - u**2)**((n - 1)/2) * u**m
            uu = cos(a*x)

        #  n      m       u=S   n         (m-1)/2
        # S(x) * C(x) dx  -->  u  * (1-u^2)       du
        elif m_:
            ff = u**n * (1 - u**2)**((m - 1)/2)
            uu = sin(a*x)

        fi = integrate(ff, u)  # XXX cyclic deps
        fx = fi.subs(u, uu)
        if conds == 'piecewise':
            return Piecewise((fx / a, Ne(a, 0)), (zz, True))
        return fx / a

    # n & m are both even
    #
    #               2k      2m                         2l       2l
    # we transform S (x) * C (x) into terms with only S (x) or C (x)
    #
    # example:
    #  100     4       100        2    2    100          4         2
    # S (x) * C (x) = S (x) * (1-S (x))  = S (x) * (1 + S (x) - 2*S (x))
    #
    #                  104       102     100
    #               = S (x) - 2*S (x) + S (x)
    #       2k
    # then S   is integrated with recursive formula

    # take largest n or m -- to choose simplest substitution
    n_ = (Abs(n) > Abs(m))
    m_ = (Abs(m) > Abs(n))
    res = S.Zero

    if n_:
        #  2k         2 k             i             2i
        # C   = (1 - S )  = sum(i, (-) * B(k, i) * S  )
        if m > 0:
            for i in range(0, m//2 + 1):
                res += (S.NegativeOne**i * binomial(m//2, i) *
                        _sin_pow_integrate(n + 2*i, x))

        elif m == 0:
            res = _sin_pow_integrate(n, x)
        else:

            # m < 0 , |n| > |m|
            #  /
            # |
            # |    m       n
            # | cos (x) sin (x) dx =
            # |
            # |
            #/
            #                                      /
            #                                     |
            #   -1        m+1     n-1     n - 1   |     m+2     n-2
            # ________ cos (x) sin (x) + _______  |  cos (x) sin (x) dx
            #                                     |
            #   m + 1                     m + 1   |
            #                                    /

            res = (Rational(-1, m + 1) * cos(x)**(m + 1) * sin(x)**(n - 1) +
                   Rational(n - 1, m + 1) *
                   trigintegrate(cos(x)**(m + 2)*sin(x)**(n - 2), x))

    elif m_:
        #  2k         2 k            i             2i
        # S   = (1 - C ) = sum(i, (-) * B(k, i) * C  )
        if n > 0:

            #      /                            /
            #     |                            |
            #     |    m       n               |    -m         n
            #     | cos (x)*sin (x) dx  or     | cos (x) * sin (x) dx
            #     |                            |
            #    /                            /
            #
            #    |m| > |n| ; m, n >0 ; m, n belong to Z - {0}
            #       n                                         2
            #    sin (x) term is expanded here in terms of cos (x),
            #    and then integrated.
            #

            for i in range(0, n//2 + 1):
                res += (S.NegativeOne**i * binomial(n//2, i) *
                        _cos_pow_integrate(m + 2*i, x))

        elif n == 0:

            #   /
            #  |
            #  |  1
            #  | _ _ _
            #  |    m
            #  | cos (x)
            # /
            #

            res = _cos_pow_integrate(m, x)
        else:

            # n < 0 , |m| > |n|
            #  /
            # |
            # |    m       n
            # | cos (x) sin (x) dx =
            # |
            # |
            #/
            #                                      /
            #                                     |
            #    1        m-1     n+1     m - 1   |     m-2     n+2
            #  _______ cos (x) sin (x) + _______  |  cos (x) sin (x) dx
            #                                     |
            #   n + 1                     n + 1   |
            #                                    /

            res = (Rational(1, n + 1) * cos(x)**(m - 1)*sin(x)**(n + 1) +
                   Rational(m - 1, n + 1) *
                   trigintegrate(cos(x)**(m - 2)*sin(x)**(n + 2), x))

    else:
        if m == n:
            ##Substitute sin(2x)/2 for sin(x)cos(x) and then Integrate.
            res = integrate((sin(2*x)*S.Half)**m, x)
        elif (m == -n):
            if n < 0:
                # Same as the scheme described above.
                # the function argument to integrate in the end will
                # be 1, this cannot be integrated by trigintegrate.
                # Hence use sympy.integrals.integrate.
                res = (Rational(1, n + 1) * cos(x)**(m - 1) * sin(x)**(n + 1) +
                       Rational(m - 1, n + 1) *
                       integrate(cos(x)**(m - 2) * sin(x)**(n + 2), x))
            else:
                res = (Rational(-1, m + 1) * cos(x)**(m + 1) * sin(x)**(n - 1) +
                       Rational(n - 1, m + 1) *
                       integrate(cos(x)**(m + 2)*sin(x)**(n - 2), x))
    if conds == 'piecewise':
        return Piecewise((res.subs(x, a*x) / a, Ne(a, 0)), (zz, True))
    return res.subs(x, a*x) / a


def _sin_pow_integrate(n, x):
    if n > 0:
        if n == 1:
            #Recursion break
            return -cos(x)

        # n > 0
        #  /                                                 /
        # |                                                 |
        # |    n           -1               n-1     n - 1   |     n-2
        # | sin (x) dx =  ______ cos (x) sin (x) + _______  |  sin (x) dx
        # |                                                 |
        # |                 n                         n     |
        #/                                                 /
        #
        #

        return (Rational(-1, n) * cos(x) * sin(x)**(n - 1) +
                Rational(n - 1, n) * _sin_pow_integrate(n - 2, x))

    if n < 0:
        if n == -1:
            ##Make sure this does not come back here again.
            ##Recursion breaks here or at n==0.
            return trigintegrate(1/sin(x), x)

        # n < 0
        #  /                                                 /
        # |                                                 |
        # |    n            1               n+1     n + 2   |     n+2
        # | sin (x) dx = _______ cos (x) sin (x) + _______  |  sin (x) dx
        # |                                                 |
        # |               n + 1                     n + 1   |
        #/                                                 /
        #

        return (Rational(1, n + 1) * cos(x) * sin(x)**(n + 1) +
                Rational(n + 2, n + 1) * _sin_pow_integrate(n + 2, x))

    else:
        #n == 0
        #Recursion break.
        return x


def _cos_pow_integrate(n, x):
    if n > 0:
        if n == 1:
            #Recursion break.
            return sin(x)

        # n > 0
        #  /                                                 /
        # |                                                 |
        # |    n            1               n-1     n - 1   |     n-2
        # | sin (x) dx =  ______ sin (x) cos (x) + _______  |  cos (x) dx
        # |                                                 |
        # |                 n                         n     |
        #/                                                 /
        #

        return (Rational(1, n) * sin(x) * cos(x)**(n - 1) +
                Rational(n - 1, n) * _cos_pow_integrate(n - 2, x))

    if n < 0:
        if n == -1:
            ##Recursion break
            return trigintegrate(1/cos(x), x)

        # n < 0
        #  /                                                 /
        # |                                                 |
        # |    n            -1              n+1     n + 2   |     n+2
        # | cos (x) dx = _______ sin (x) cos (x) + _______  |  cos (x) dx
        # |                                                 |
        # |               n + 1                     n + 1   |
        #/                                                 /
        #

        return (Rational(-1, n + 1) * sin(x) * cos(x)**(n + 1) +
                Rational(n + 2, n + 1) * _cos_pow_integrate(n + 2, x))
    else:
        # n == 0
        #Recursion Break.
        return x

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\vector\operators.py ===
import collections
from sympy.core.expr import Expr
from sympy.core import sympify, S, preorder_traversal
from sympy.vector.coordsysrect import CoordSys3D
from sympy.vector.vector import Vector, VectorMul, VectorAdd, Cross, Dot
from sympy.core.function import Derivative
from sympy.core.add import Add
from sympy.core.mul import Mul


def _get_coord_systems(expr):
    g = preorder_traversal(expr)
    ret = set()
    for i in g:
        if isinstance(i, CoordSys3D):
            ret.add(i)
            g.skip()
    return frozenset(ret)


def _split_mul_args_wrt_coordsys(expr):
    d = collections.defaultdict(lambda: S.One)
    for i in expr.args:
        d[_get_coord_systems(i)] *= i
    return list(d.values())


class Gradient(Expr):
    """
    Represents unevaluated Gradient.

    Examples
    ========

    >>> from sympy.vector import CoordSys3D, Gradient
    >>> R = CoordSys3D('R')
    >>> s = R.x*R.y*R.z
    >>> Gradient(s)
    Gradient(R.x*R.y*R.z)

    """

    def __new__(cls, expr):
        expr = sympify(expr)
        obj = Expr.__new__(cls, expr)
        obj._expr = expr
        return obj

    def doit(self, **hints):
        return gradient(self._expr, doit=True)


class Divergence(Expr):
    """
    Represents unevaluated Divergence.

    Examples
    ========

    >>> from sympy.vector import CoordSys3D, Divergence
    >>> R = CoordSys3D('R')
    >>> v = R.y*R.z*R.i + R.x*R.z*R.j + R.x*R.y*R.k
    >>> Divergence(v)
    Divergence(R.y*R.z*R.i + R.x*R.z*R.j + R.x*R.y*R.k)

    """

    def __new__(cls, expr):
        expr = sympify(expr)
        obj = Expr.__new__(cls, expr)
        obj._expr = expr
        return obj

    def doit(self, **hints):
        return divergence(self._expr, doit=True)


class Curl(Expr):
    """
    Represents unevaluated Curl.

    Examples
    ========

    >>> from sympy.vector import CoordSys3D, Curl
    >>> R = CoordSys3D('R')
    >>> v = R.y*R.z*R.i + R.x*R.z*R.j + R.x*R.y*R.k
    >>> Curl(v)
    Curl(R.y*R.z*R.i + R.x*R.z*R.j + R.x*R.y*R.k)

    """

    def __new__(cls, expr):
        expr = sympify(expr)
        obj = Expr.__new__(cls, expr)
        obj._expr = expr
        return obj

    def doit(self, **hints):
        return curl(self._expr, doit=True)


def curl(vect, doit=True):
    """
    Returns the curl of a vector field computed wrt the base scalars
    of the given coordinate system.

    Parameters
    ==========

    vect : Vector
        The vector operand

    doit : bool
        If True, the result is returned after calling .doit() on
        each component. Else, the returned expression contains
        Derivative instances

    Examples
    ========

    >>> from sympy.vector import CoordSys3D, curl
    >>> R = CoordSys3D('R')
    >>> v1 = R.y*R.z*R.i + R.x*R.z*R.j + R.x*R.y*R.k
    >>> curl(v1)
    0
    >>> v2 = R.x*R.y*R.z*R.i
    >>> curl(v2)
    R.x*R.y*R.j + (-R.x*R.z)*R.k

    """

    coord_sys = _get_coord_systems(vect)

    if len(coord_sys) == 0:
        return Vector.zero
    elif len(coord_sys) == 1:
        coord_sys = next(iter(coord_sys))
        i, j, k = coord_sys.base_vectors()
        x, y, z = coord_sys.base_scalars()
        h1, h2, h3 = coord_sys.lame_coefficients()
        vectx = vect.dot(i)
        vecty = vect.dot(j)
        vectz = vect.dot(k)
        outvec = Vector.zero
        outvec += (Derivative(vectz * h3, y) -
                   Derivative(vecty * h2, z)) * i / (h2 * h3)
        outvec += (Derivative(vectx * h1, z) -
                   Derivative(vectz * h3, x)) * j / (h1 * h3)
        outvec += (Derivative(vecty * h2, x) -
                   Derivative(vectx * h1, y)) * k / (h2 * h1)

        if doit:
            return outvec.doit()
        return outvec
    else:
        if isinstance(vect, (Add, VectorAdd)):
            from sympy.vector import express
            try:
                cs = next(iter(coord_sys))
                args = [express(i, cs, variables=True) for i in vect.args]
            except ValueError:
                args = vect.args
            return VectorAdd.fromiter(curl(i, doit=doit) for i in args)
        elif isinstance(vect, (Mul, VectorMul)):
            vector = [i for i in vect.args if isinstance(i, (Vector, Cross, Gradient))][0]
            scalar = Mul.fromiter(i for i in vect.args if not isinstance(i, (Vector, Cross, Gradient)))
            res = Cross(gradient(scalar), vector).doit() + scalar*curl(vector, doit=doit)
            if doit:
                return res.doit()
            return res
        elif isinstance(vect, (Cross, Curl, Gradient)):
            return Curl(vect)
        else:
            raise ValueError("Invalid argument for curl")


def divergence(vect, doit=True):
    """
    Returns the divergence of a vector field computed wrt the base
    scalars of the given coordinate system.

    Parameters
    ==========

    vector : Vector
        The vector operand

    doit : bool
        If True, the result is returned after calling .doit() on
        each component. Else, the returned expression contains
        Derivative instances

    Examples
    ========

    >>> from sympy.vector import CoordSys3D, divergence
    >>> R = CoordSys3D('R')
    >>> v1 = R.x*R.y*R.z * (R.i+R.j+R.k)

    >>> divergence(v1)
    R.x*R.y + R.x*R.z + R.y*R.z
    >>> v2 = 2*R.y*R.z*R.j
    >>> divergence(v2)
    2*R.z

    """
    coord_sys = _get_coord_systems(vect)
    if len(coord_sys) == 0:
        return S.Zero
    elif len(coord_sys) == 1:
        if isinstance(vect, (Cross, Curl, Gradient)):
            return Divergence(vect)
        # TODO: is case of many coord systems, this gets a random one:
        coord_sys = next(iter(coord_sys))
        i, j, k = coord_sys.base_vectors()
        x, y, z = coord_sys.base_scalars()
        h1, h2, h3 = coord_sys.lame_coefficients()
        vx = _diff_conditional(vect.dot(i), x, h2, h3) \
             / (h1 * h2 * h3)
        vy = _diff_conditional(vect.dot(j), y, h3, h1) \
             / (h1 * h2 * h3)
        vz = _diff_conditional(vect.dot(k), z, h1, h2) \
             / (h1 * h2 * h3)
        res = vx + vy + vz
        if doit:
            return res.doit()
        return res
    else:
        if isinstance(vect, (Add, VectorAdd)):
            return Add.fromiter(divergence(i, doit=doit) for i in vect.args)
        elif isinstance(vect, (Mul, VectorMul)):
            vector = [i for i in vect.args if isinstance(i, (Vector, Cross, Gradient))][0]
            scalar = Mul.fromiter(i for i in vect.args if not isinstance(i, (Vector, Cross, Gradient)))
            res = Dot(vector, gradient(scalar)) + scalar*divergence(vector, doit=doit)
            if doit:
                return res.doit()
            return res
        elif isinstance(vect, (Cross, Curl, Gradient)):
            return Divergence(vect)
        else:
            raise ValueError("Invalid argument for divergence")


def gradient(scalar_field, doit=True):
    """
    Returns the vector gradient of a scalar field computed wrt the
    base scalars of the given coordinate system.

    Parameters
    ==========

    scalar_field : SymPy Expr
        The scalar field to compute the gradient of

    doit : bool
        If True, the result is returned after calling .doit() on
        each component. Else, the returned expression contains
        Derivative instances

    Examples
    ========

    >>> from sympy.vector import CoordSys3D, gradient
    >>> R = CoordSys3D('R')
    >>> s1 = R.x*R.y*R.z
    >>> gradient(s1)
    R.y*R.z*R.i + R.x*R.z*R.j + R.x*R.y*R.k
    >>> s2 = 5*R.x**2*R.z
    >>> gradient(s2)
    10*R.x*R.z*R.i + 5*R.x**2*R.k

    """
    coord_sys = _get_coord_systems(scalar_field)

    if len(coord_sys) == 0:
        return Vector.zero
    elif len(coord_sys) == 1:
        coord_sys = next(iter(coord_sys))
        h1, h2, h3 = coord_sys.lame_coefficients()
        i, j, k = coord_sys.base_vectors()
        x, y, z = coord_sys.base_scalars()
        vx = Derivative(scalar_field, x) / h1
        vy = Derivative(scalar_field, y) / h2
        vz = Derivative(scalar_field, z) / h3

        if doit:
            return (vx * i + vy * j + vz * k).doit()
        return vx * i + vy * j + vz * k
    else:
        if isinstance(scalar_field, (Add, VectorAdd)):
            return VectorAdd.fromiter(gradient(i) for i in scalar_field.args)
        if isinstance(scalar_field, (Mul, VectorMul)):
            s = _split_mul_args_wrt_coordsys(scalar_field)
            return VectorAdd.fromiter(scalar_field / i * gradient(i) for i in s)
        return Gradient(scalar_field)


class Laplacian(Expr):
    """
    Represents unevaluated Laplacian.

    Examples
    ========

    >>> from sympy.vector import CoordSys3D, Laplacian
    >>> R = CoordSys3D('R')
    >>> v = 3*R.x**3*R.y**2*R.z**3
    >>> Laplacian(v)
    Laplacian(3*R.x**3*R.y**2*R.z**3)

    """

    def __new__(cls, expr):
        expr = sympify(expr)
        obj = Expr.__new__(cls, expr)
        obj._expr = expr
        return obj

    def doit(self, **hints):
        from sympy.vector.functions import laplacian
        return laplacian(self._expr)


def _diff_conditional(expr, base_scalar, coeff_1, coeff_2):
    """
    First re-expresses expr in the system that base_scalar belongs to.
    If base_scalar appears in the re-expressed form, differentiates
    it wrt base_scalar.
    Else, returns 0
    """
    from sympy.vector.functions import express
    new_expr = express(expr, base_scalar.system, variables=True)
    arg = coeff_1 * coeff_2 * new_expr
    return Derivative(arg, base_scalar) if arg else S.Zero

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\urllib3\exceptions.py ===
from __future__ import annotations

import socket
import typing
import warnings
from email.errors import MessageDefect
from http.client import IncompleteRead as httplib_IncompleteRead

if typing.TYPE_CHECKING:
    from .connection import HTTPConnection
    from .connectionpool import ConnectionPool
    from .response import HTTPResponse
    from .util.retry import Retry

# Base Exceptions


class HTTPError(Exception):
    """Base exception used by this module."""


class HTTPWarning(Warning):
    """Base warning used by this module."""


_TYPE_REDUCE_RESULT = tuple[typing.Callable[..., object], tuple[object, ...]]


class PoolError(HTTPError):
    """Base exception for errors caused within a pool."""

    def __init__(self, pool: ConnectionPool, message: str) -> None:
        self.pool = pool
        self._message = message
        super().__init__(f"{pool}: {message}")

    def __reduce__(self) -> _TYPE_REDUCE_RESULT:
        # For pickling purposes.
        return self.__class__, (None, self._message)


class RequestError(PoolError):
    """Base exception for PoolErrors that have associated URLs."""

    def __init__(self, pool: ConnectionPool, url: str, message: str) -> None:
        self.url = url
        super().__init__(pool, message)

    def __reduce__(self) -> _TYPE_REDUCE_RESULT:
        # For pickling purposes.
        return self.__class__, (None, self.url, self._message)


class SSLError(HTTPError):
    """Raised when SSL certificate fails in an HTTPS connection."""


class ProxyError(HTTPError):
    """Raised when the connection to a proxy fails."""

    # The original error is also available as __cause__.
    original_error: Exception

    def __init__(self, message: str, error: Exception) -> None:
        super().__init__(message, error)
        self.original_error = error


class DecodeError(HTTPError):
    """Raised when automatic decoding based on Content-Type fails."""


class ProtocolError(HTTPError):
    """Raised when something unexpected happens mid-request/response."""


#: Renamed to ProtocolError but aliased for backwards compatibility.
ConnectionError = ProtocolError


# Leaf Exceptions


class MaxRetryError(RequestError):
    """Raised when the maximum number of retries is exceeded.

    :param pool: The connection pool
    :type pool: :class:`~urllib3.connectionpool.HTTPConnectionPool`
    :param str url: The requested Url
    :param reason: The underlying error
    :type reason: :class:`Exception`

    """

    def __init__(
        self, pool: ConnectionPool, url: str, reason: Exception | None = None
    ) -> None:
        self.reason = reason

        message = f"Max retries exceeded with url: {url} (Caused by {reason!r})"

        super().__init__(pool, url, message)

    def __reduce__(self) -> _TYPE_REDUCE_RESULT:
        # For pickling purposes.
        return self.__class__, (None, self.url, self.reason)


class HostChangedError(RequestError):
    """Raised when an existing pool gets a request for a foreign host."""

    def __init__(
        self, pool: ConnectionPool, url: str, retries: Retry | int = 3
    ) -> None:
        message = f"Tried to open a foreign host with url: {url}"
        super().__init__(pool, url, message)
        self.retries = retries


class TimeoutStateError(HTTPError):
    """Raised when passing an invalid state to a timeout"""


class TimeoutError(HTTPError):
    """Raised when a socket timeout error occurs.

    Catching this error will catch both :exc:`ReadTimeoutErrors
    <ReadTimeoutError>` and :exc:`ConnectTimeoutErrors <ConnectTimeoutError>`.
    """


class ReadTimeoutError(TimeoutError, RequestError):
    """Raised when a socket timeout occurs while receiving data from a server"""


# This timeout error does not have a URL attached and needs to inherit from the
# base HTTPError
class ConnectTimeoutError(TimeoutError):
    """Raised when a socket timeout occurs while connecting to a server"""


class NewConnectionError(ConnectTimeoutError, HTTPError):
    """Raised when we fail to establish a new connection. Usually ECONNREFUSED."""

    def __init__(self, conn: HTTPConnection, message: str) -> None:
        self.conn = conn
        self._message = message
        super().__init__(f"{conn}: {message}")

    def __reduce__(self) -> _TYPE_REDUCE_RESULT:
        # For pickling purposes.
        return self.__class__, (None, self._message)

    @property
    def pool(self) -> HTTPConnection:
        warnings.warn(
            "The 'pool' property is deprecated and will be removed "
            "in urllib3 v2.1.0. Use 'conn' instead.",
            DeprecationWarning,
            stacklevel=2,
        )

        return self.conn


class NameResolutionError(NewConnectionError):
    """Raised when host name resolution fails."""

    def __init__(self, host: str, conn: HTTPConnection, reason: socket.gaierror):
        message = f"Failed to resolve '{host}' ({reason})"
        self._host = host
        self._reason = reason
        super().__init__(conn, message)

    def __reduce__(self) -> _TYPE_REDUCE_RESULT:
        # For pickling purposes.
        return self.__class__, (self._host, None, self._reason)


class EmptyPoolError(PoolError):
    """Raised when a pool runs out of connections and no more are allowed."""


class FullPoolError(PoolError):
    """Raised when we try to add a connection to a full pool in blocking mode."""


class ClosedPoolError(PoolError):
    """Raised when a request enters a pool after the pool has been closed."""


class LocationValueError(ValueError, HTTPError):
    """Raised when there is something wrong with a given URL input."""


class LocationParseError(LocationValueError):
    """Raised when get_host or similar fails to parse the URL input."""

    def __init__(self, location: str) -> None:
        message = f"Failed to parse: {location}"
        super().__init__(message)

        self.location = location


class URLSchemeUnknown(LocationValueError):
    """Raised when a URL input has an unsupported scheme."""

    def __init__(self, scheme: str):
        message = f"Not supported URL scheme {scheme}"
        super().__init__(message)

        self.scheme = scheme


class ResponseError(HTTPError):
    """Used as a container for an error reason supplied in a MaxRetryError."""

    GENERIC_ERROR = "too many error responses"
    SPECIFIC_ERROR = "too many {status_code} error responses"


class SecurityWarning(HTTPWarning):
    """Warned when performing security reducing actions"""


class InsecureRequestWarning(SecurityWarning):
    """Warned when making an unverified HTTPS request."""


class NotOpenSSLWarning(SecurityWarning):
    """Warned when using unsupported SSL library"""


class SystemTimeWarning(SecurityWarning):
    """Warned when system time is suspected to be wrong"""


class InsecurePlatformWarning(SecurityWarning):
    """Warned when certain TLS/SSL configuration is not available on a platform."""


class DependencyWarning(HTTPWarning):
    """
    Warned when an attempt is made to import a module with missing optional
    dependencies.
    """


class ResponseNotChunked(ProtocolError, ValueError):
    """Response needs to be chunked in order to read it as chunks."""


class BodyNotHttplibCompatible(HTTPError):
    """
    Body should be :class:`http.client.HTTPResponse` like
    (have an fp attribute which returns raw chunks) for read_chunked().
    """


class IncompleteRead(HTTPError, httplib_IncompleteRead):
    """
    Response length doesn't match expected Content-Length

    Subclass of :class:`http.client.IncompleteRead` to allow int value
    for ``partial`` to avoid creating large objects on streamed reads.
    """

    partial: int  # type: ignore[assignment]
    expected: int

    def __init__(self, partial: int, expected: int) -> None:
        self.partial = partial
        self.expected = expected

    def __repr__(self) -> str:
        return "IncompleteRead(%i bytes read, %i more expected)" % (
            self.partial,
            self.expected,
        )


class InvalidChunkLength(HTTPError, httplib_IncompleteRead):
    """Invalid chunk length in a chunked response."""

    def __init__(self, response: HTTPResponse, length: bytes) -> None:
        self.partial: int = response.tell()  # type: ignore[assignment]
        self.expected: int | None = response.length_remaining
        self.response = response
        self.length = length

    def __repr__(self) -> str:
        return "InvalidChunkLength(got length %r, %i bytes read)" % (
            self.length,
            self.partial,
        )


class InvalidHeader(HTTPError):
    """The header provided was somehow invalid."""


class ProxySchemeUnknown(AssertionError, URLSchemeUnknown):
    """ProxyManager does not support the supplied scheme"""

    # TODO(t-8ch): Stop inheriting from AssertionError in v2.0.

    def __init__(self, scheme: str | None) -> None:
        # 'localhost' is here because our URL parser parses
        # localhost:8080 -> scheme=localhost, remove if we fix this.
        if scheme == "localhost":
            scheme = None
        if scheme is None:
            message = "Proxy URL had no scheme, should start with http:// or https://"
        else:
            message = f"Proxy URL had unsupported scheme {scheme}, should use http:// or https://"
        super().__init__(message)


class ProxySchemeUnsupported(ValueError):
    """Fetching HTTPS resources through HTTPS proxies is unsupported"""


class HeaderParsingError(HTTPError):
    """Raised by assert_header_parsing, but we convert it to a log.warning statement."""

    def __init__(
        self, defects: list[MessageDefect], unparsed_data: bytes | str | None
    ) -> None:
        message = f"{defects or 'Unknown'}, unparsed data: {unparsed_data!r}"
        super().__init__(message)


class UnrewindableBodyError(HTTPError):
    """urllib3 encountered an error when trying to rewind a body"""

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\io\excel\_util.py ===
from __future__ import annotations

from collections.abc import (
    Hashable,
    Iterable,
    MutableMapping,
    Sequence,
)
from typing import (
    TYPE_CHECKING,
    Any,
    Callable,
    Literal,
    TypeVar,
    overload,
)

from pandas.compat._optional import import_optional_dependency

from pandas.core.dtypes.common import (
    is_integer,
    is_list_like,
)

if TYPE_CHECKING:
    from pandas.io.excel._base import ExcelWriter

    ExcelWriter_t = type[ExcelWriter]
    usecols_func = TypeVar("usecols_func", bound=Callable[[Hashable], object])

_writers: MutableMapping[str, ExcelWriter_t] = {}


def register_writer(klass: ExcelWriter_t) -> None:
    """
    Add engine to the excel writer registry.io.excel.

    You must use this method to integrate with ``to_excel``.

    Parameters
    ----------
    klass : ExcelWriter
    """
    if not callable(klass):
        raise ValueError("Can only register callables as engines")
    engine_name = klass._engine
    _writers[engine_name] = klass


def get_default_engine(ext: str, mode: Literal["reader", "writer"] = "reader") -> str:
    """
    Return the default reader/writer for the given extension.

    Parameters
    ----------
    ext : str
        The excel file extension for which to get the default engine.
    mode : str {'reader', 'writer'}
        Whether to get the default engine for reading or writing.
        Either 'reader' or 'writer'

    Returns
    -------
    str
        The default engine for the extension.
    """
    _default_readers = {
        "xlsx": "openpyxl",
        "xlsm": "openpyxl",
        "xlsb": "pyxlsb",
        "xls": "xlrd",
        "ods": "odf",
    }
    _default_writers = {
        "xlsx": "openpyxl",
        "xlsm": "openpyxl",
        "xlsb": "pyxlsb",
        "ods": "odf",
    }
    assert mode in ["reader", "writer"]
    if mode == "writer":
        # Prefer xlsxwriter over openpyxl if installed
        xlsxwriter = import_optional_dependency("xlsxwriter", errors="warn")
        if xlsxwriter:
            _default_writers["xlsx"] = "xlsxwriter"
        return _default_writers[ext]
    else:
        return _default_readers[ext]


def get_writer(engine_name: str) -> ExcelWriter_t:
    try:
        return _writers[engine_name]
    except KeyError as err:
        raise ValueError(f"No Excel writer '{engine_name}'") from err


def _excel2num(x: str) -> int:
    """
    Convert Excel column name like 'AB' to 0-based column index.

    Parameters
    ----------
    x : str
        The Excel column name to convert to a 0-based column index.

    Returns
    -------
    num : int
        The column index corresponding to the name.

    Raises
    ------
    ValueError
        Part of the Excel column name was invalid.
    """
    index = 0

    for c in x.upper().strip():
        cp = ord(c)

        if cp < ord("A") or cp > ord("Z"):
            raise ValueError(f"Invalid column name: {x}")

        index = index * 26 + cp - ord("A") + 1

    return index - 1


def _range2cols(areas: str) -> list[int]:
    """
    Convert comma separated list of column names and ranges to indices.

    Parameters
    ----------
    areas : str
        A string containing a sequence of column ranges (or areas).

    Returns
    -------
    cols : list
        A list of 0-based column indices.

    Examples
    --------
    >>> _range2cols('A:E')
    [0, 1, 2, 3, 4]
    >>> _range2cols('A,C,Z:AB')
    [0, 2, 25, 26, 27]
    """
    cols: list[int] = []

    for rng in areas.split(","):
        if ":" in rng:
            rngs = rng.split(":")
            cols.extend(range(_excel2num(rngs[0]), _excel2num(rngs[1]) + 1))
        else:
            cols.append(_excel2num(rng))

    return cols


@overload
def maybe_convert_usecols(usecols: str | list[int]) -> list[int]:
    ...


@overload
def maybe_convert_usecols(usecols: list[str]) -> list[str]:
    ...


@overload
def maybe_convert_usecols(usecols: usecols_func) -> usecols_func:
    ...


@overload
def maybe_convert_usecols(usecols: None) -> None:
    ...


def maybe_convert_usecols(
    usecols: str | list[int] | list[str] | usecols_func | None,
) -> None | list[int] | list[str] | usecols_func:
    """
    Convert `usecols` into a compatible format for parsing in `parsers.py`.

    Parameters
    ----------
    usecols : object
        The use-columns object to potentially convert.

    Returns
    -------
    converted : object
        The compatible format of `usecols`.
    """
    if usecols is None:
        return usecols

    if is_integer(usecols):
        raise ValueError(
            "Passing an integer for `usecols` is no longer supported.  "
            "Please pass in a list of int from 0 to `usecols` inclusive instead."
        )

    if isinstance(usecols, str):
        return _range2cols(usecols)

    return usecols


@overload
def validate_freeze_panes(freeze_panes: tuple[int, int]) -> Literal[True]:
    ...


@overload
def validate_freeze_panes(freeze_panes: None) -> Literal[False]:
    ...


def validate_freeze_panes(freeze_panes: tuple[int, int] | None) -> bool:
    if freeze_panes is not None:
        if len(freeze_panes) == 2 and all(
            isinstance(item, int) for item in freeze_panes
        ):
            return True

        raise ValueError(
            "freeze_panes must be of form (row, column) "
            "where row and column are integers"
        )

    # freeze_panes wasn't specified, return False so it won't be applied
    # to output sheet
    return False


def fill_mi_header(
    row: list[Hashable], control_row: list[bool]
) -> tuple[list[Hashable], list[bool]]:
    """
    Forward fill blank entries in row but only inside the same parent index.

    Used for creating headers in Multiindex.

    Parameters
    ----------
    row : list
        List of items in a single row.
    control_row : list of bool
        Helps to determine if particular column is in same parent index as the
        previous value. Used to stop propagation of empty cells between
        different indexes.

    Returns
    -------
    Returns changed row and control_row
    """
    last = row[0]
    for i in range(1, len(row)):
        if not control_row[i]:
            last = row[i]

        if row[i] == "" or row[i] is None:
            row[i] = last
        else:
            control_row[i] = False
            last = row[i]

    return row, control_row


def pop_header_name(
    row: list[Hashable], index_col: int | Sequence[int]
) -> tuple[Hashable | None, list[Hashable]]:
    """
    Pop the header name for MultiIndex parsing.

    Parameters
    ----------
    row : list
        The data row to parse for the header name.
    index_col : int, list
        The index columns for our data. Assumed to be non-null.

    Returns
    -------
    header_name : str
        The extracted header name.
    trimmed_row : list
        The original data row with the header name removed.
    """
    # Pop out header name and fill w/blank.
    if is_list_like(index_col):
        assert isinstance(index_col, Iterable)
        i = max(index_col)
    else:
        assert not isinstance(index_col, Iterable)
        i = index_col

    header_name = row[i]
    header_name = None if header_name == "" else header_name

    return header_name, row[:i] + [""] + row[i + 1 :]


def combine_kwargs(engine_kwargs: dict[str, Any] | None, kwargs: dict) -> dict:
    """
    Used to combine two sources of kwargs for the backend engine.

    Use of kwargs is deprecated, this function is solely for use in 1.3 and should
    be removed in 1.4/2.0. Also _base.ExcelWriter.__new__ ensures either engine_kwargs
    or kwargs must be None or empty respectively.

    Parameters
    ----------
    engine_kwargs: dict
        kwargs to be passed through to the engine.
    kwargs: dict
        kwargs to be psased through to the engine (deprecated)

    Returns
    -------
    engine_kwargs combined with kwargs
    """
    if engine_kwargs is None:
        result = {}
    else:
        result = engine_kwargs.copy()
    result.update(kwargs)
    return result

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\functions\special\spherical_harmonics.py ===
from sympy.core.expr import Expr
from sympy.core.function import DefinedFunction, ArgumentIndexError
from sympy.core.numbers import I, pi
from sympy.core.singleton import S
from sympy.core.symbol import Dummy
from sympy.functions import assoc_legendre
from sympy.functions.combinatorial.factorials import factorial
from sympy.functions.elementary.complexes import Abs, conjugate
from sympy.functions.elementary.exponential import exp
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.trigonometric import sin, cos, cot

_x = Dummy("x")

class Ynm(DefinedFunction):
    r"""
    Spherical harmonics defined as

    .. math::
        Y_n^m(\theta, \varphi) := \sqrt{\frac{(2n+1)(n-m)!}{4\pi(n+m)!}}
                                  \exp(i m \varphi)
                                  \mathrm{P}_n^m\left(\cos(\theta)\right)

    Explanation
    ===========

    ``Ynm()`` gives the spherical harmonic function of order $n$ and $m$
    in $\theta$ and $\varphi$, $Y_n^m(\theta, \varphi)$. The four
    parameters are as follows: $n \geq 0$ an integer and $m$ an integer
    such that $-n \leq m \leq n$ holds. The two angles are real-valued
    with $\theta \in [0, \pi]$ and $\varphi \in [0, 2\pi]$.

    Examples
    ========

    >>> from sympy import Ynm, Symbol, simplify
    >>> from sympy.abc import n,m
    >>> theta = Symbol("theta")
    >>> phi = Symbol("phi")

    >>> Ynm(n, m, theta, phi)
    Ynm(n, m, theta, phi)

    Several symmetries are known, for the order:

    >>> Ynm(n, -m, theta, phi)
    (-1)**m*exp(-2*I*m*phi)*Ynm(n, m, theta, phi)

    As well as for the angles:

    >>> Ynm(n, m, -theta, phi)
    Ynm(n, m, theta, phi)

    >>> Ynm(n, m, theta, -phi)
    exp(-2*I*m*phi)*Ynm(n, m, theta, phi)

    For specific integers $n$ and $m$ we can evaluate the harmonics
    to more useful expressions:

    >>> simplify(Ynm(0, 0, theta, phi).expand(func=True))
    1/(2*sqrt(pi))

    >>> simplify(Ynm(1, -1, theta, phi).expand(func=True))
    sqrt(6)*exp(-I*phi)*sin(theta)/(4*sqrt(pi))

    >>> simplify(Ynm(1, 0, theta, phi).expand(func=True))
    sqrt(3)*cos(theta)/(2*sqrt(pi))

    >>> simplify(Ynm(1, 1, theta, phi).expand(func=True))
    -sqrt(6)*exp(I*phi)*sin(theta)/(4*sqrt(pi))

    >>> simplify(Ynm(2, -2, theta, phi).expand(func=True))
    sqrt(30)*exp(-2*I*phi)*sin(theta)**2/(8*sqrt(pi))

    >>> simplify(Ynm(2, -1, theta, phi).expand(func=True))
    sqrt(30)*exp(-I*phi)*sin(2*theta)/(8*sqrt(pi))

    >>> simplify(Ynm(2, 0, theta, phi).expand(func=True))
    sqrt(5)*(3*cos(theta)**2 - 1)/(4*sqrt(pi))

    >>> simplify(Ynm(2, 1, theta, phi).expand(func=True))
    -sqrt(30)*exp(I*phi)*sin(2*theta)/(8*sqrt(pi))

    >>> simplify(Ynm(2, 2, theta, phi).expand(func=True))
    sqrt(30)*exp(2*I*phi)*sin(theta)**2/(8*sqrt(pi))

    We can differentiate the functions with respect
    to both angles:

    >>> from sympy import Ynm, Symbol, diff
    >>> from sympy.abc import n,m
    >>> theta = Symbol("theta")
    >>> phi = Symbol("phi")

    >>> diff(Ynm(n, m, theta, phi), theta)
    m*cot(theta)*Ynm(n, m, theta, phi) + sqrt((-m + n)*(m + n + 1))*exp(-I*phi)*Ynm(n, m + 1, theta, phi)

    >>> diff(Ynm(n, m, theta, phi), phi)
    I*m*Ynm(n, m, theta, phi)

    Further we can compute the complex conjugation:

    >>> from sympy import Ynm, Symbol, conjugate
    >>> from sympy.abc import n,m
    >>> theta = Symbol("theta")
    >>> phi = Symbol("phi")

    >>> conjugate(Ynm(n, m, theta, phi))
    (-1)**(2*m)*exp(-2*I*m*phi)*Ynm(n, m, theta, phi)

    To get back the well known expressions in spherical
    coordinates, we use full expansion:

    >>> from sympy import Ynm, Symbol, expand_func
    >>> from sympy.abc import n,m
    >>> theta = Symbol("theta")
    >>> phi = Symbol("phi")

    >>> expand_func(Ynm(n, m, theta, phi))
    sqrt((2*n + 1)*factorial(-m + n)/factorial(m + n))*exp(I*m*phi)*assoc_legendre(n, m, cos(theta))/(2*sqrt(pi))

    See Also
    ========

    Ynm_c, Znm

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Spherical_harmonics
    .. [2] https://mathworld.wolfram.com/SphericalHarmonic.html
    .. [3] https://functions.wolfram.com/Polynomials/SphericalHarmonicY/
    .. [4] https://dlmf.nist.gov/14.30

    """

    @classmethod
    def eval(cls, n, m, theta, phi):
        # Handle negative index m and arguments theta, phi
        if m.could_extract_minus_sign():
            m = -m
            return S.NegativeOne**m * exp(-2*I*m*phi) * Ynm(n, m, theta, phi)
        if theta.could_extract_minus_sign():
            theta = -theta
            return Ynm(n, m, theta, phi)
        if phi.could_extract_minus_sign():
            phi = -phi
            return exp(-2*I*m*phi) * Ynm(n, m, theta, phi)

        # TODO Add more simplififcation here

    def _eval_expand_func(self, **hints):
        n, m, theta, phi = self.args
        rv = (sqrt((2*n + 1)/(4*pi) * factorial(n - m)/factorial(n + m)) *
                exp(I*m*phi) * assoc_legendre(n, m, cos(theta)))
        # We can do this because of the range of theta
        return rv.subs(sqrt(-cos(theta)**2 + 1), sin(theta))

    def fdiff(self, argindex=4):
        if argindex == 1:
            # Diff wrt n
            raise ArgumentIndexError(self, argindex)
        elif argindex == 2:
            # Diff wrt m
            raise ArgumentIndexError(self, argindex)
        elif argindex == 3:
            # Diff wrt theta
            n, m, theta, phi = self.args
            return (m * cot(theta) * Ynm(n, m, theta, phi) +
                    sqrt((n - m)*(n + m + 1)) * exp(-I*phi) * Ynm(n, m + 1, theta, phi))
        elif argindex == 4:
            # Diff wrt phi
            n, m, theta, phi = self.args
            return I * m * Ynm(n, m, theta, phi)
        else:
            raise ArgumentIndexError(self, argindex)

    def _eval_rewrite_as_polynomial(self, n, m, theta, phi, **kwargs):
        # TODO: Make sure n \in N
        # TODO: Assert |m| <= n ortherwise we should return 0
        return self.expand(func=True)

    def _eval_rewrite_as_sin(self, n, m, theta, phi, **kwargs):
        return self.rewrite(cos)

    def _eval_rewrite_as_cos(self, n, m, theta, phi, **kwargs):
        # This method can be expensive due to extensive use of simplification!
        from sympy.simplify import simplify, trigsimp
        # TODO: Make sure n \in N
        # TODO: Assert |m| <= n ortherwise we should return 0
        term = simplify(self.expand(func=True))
        # We can do this because of the range of theta
        term = term.xreplace({Abs(sin(theta)):sin(theta)})
        return simplify(trigsimp(term))

    def _eval_conjugate(self):
        # TODO: Make sure theta \in R and phi \in R
        n, m, theta, phi = self.args
        return S.NegativeOne**m * self.func(n, -m, theta, phi)

    def as_real_imag(self, deep=True, **hints):
        # TODO: Handle deep and hints
        n, m, theta, phi = self.args
        re = (sqrt((2*n + 1)/(4*pi) * factorial(n - m)/factorial(n + m)) *
              cos(m*phi) * assoc_legendre(n, m, cos(theta)))
        im = (sqrt((2*n + 1)/(4*pi) * factorial(n - m)/factorial(n + m)) *
              sin(m*phi) * assoc_legendre(n, m, cos(theta)))
        return (re, im)

    def _eval_evalf(self, prec):
        # Note: works without this function by just calling
        #       mpmath for Legendre polynomials. But using
        #       the dedicated function directly is cleaner.
        from mpmath import mp, workprec
        n = self.args[0]._to_mpmath(prec)
        m = self.args[1]._to_mpmath(prec)
        theta = self.args[2]._to_mpmath(prec)
        phi = self.args[3]._to_mpmath(prec)
        with workprec(prec):
            res = mp.spherharm(n, m, theta, phi)
        return Expr._from_mpmath(res, prec)


def Ynm_c(n, m, theta, phi):
    r"""
    Conjugate spherical harmonics defined as

    .. math::
        \overline{Y_n^m(\theta, \varphi)} := (-1)^m Y_n^{-m}(\theta, \varphi).

    Examples
    ========

    >>> from sympy import Ynm_c, Symbol, simplify
    >>> from sympy.abc import n,m
    >>> theta = Symbol("theta")
    >>> phi = Symbol("phi")
    >>> Ynm_c(n, m, theta, phi)
    (-1)**(2*m)*exp(-2*I*m*phi)*Ynm(n, m, theta, phi)
    >>> Ynm_c(n, m, -theta, phi)
    (-1)**(2*m)*exp(-2*I*m*phi)*Ynm(n, m, theta, phi)

    For specific integers $n$ and $m$ we can evaluate the harmonics
    to more useful expressions:

    >>> simplify(Ynm_c(0, 0, theta, phi).expand(func=True))
    1/(2*sqrt(pi))
    >>> simplify(Ynm_c(1, -1, theta, phi).expand(func=True))
    sqrt(6)*exp(I*(-phi + 2*conjugate(phi)))*sin(theta)/(4*sqrt(pi))

    See Also
    ========

    Ynm, Znm

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Spherical_harmonics
    .. [2] https://mathworld.wolfram.com/SphericalHarmonic.html
    .. [3] https://functions.wolfram.com/Polynomials/SphericalHarmonicY/

    """
    return conjugate(Ynm(n, m, theta, phi))


class Znm(DefinedFunction):
    r"""
    Real spherical harmonics defined as

    .. math::

        Z_n^m(\theta, \varphi) :=
        \begin{cases}
          \frac{Y_n^m(\theta, \varphi) + \overline{Y_n^m(\theta, \varphi)}}{\sqrt{2}} &\quad m > 0 \\
          Y_n^m(\theta, \varphi) &\quad m = 0 \\
          \frac{Y_n^m(\theta, \varphi) - \overline{Y_n^m(\theta, \varphi)}}{i \sqrt{2}} &\quad m < 0 \\
        \end{cases}

    which gives in simplified form

    .. math::

        Z_n^m(\theta, \varphi) =
        \begin{cases}
          \frac{Y_n^m(\theta, \varphi) + (-1)^m Y_n^{-m}(\theta, \varphi)}{\sqrt{2}} &\quad m > 0 \\
          Y_n^m(\theta, \varphi) &\quad m = 0 \\
          \frac{Y_n^m(\theta, \varphi) - (-1)^m Y_n^{-m}(\theta, \varphi)}{i \sqrt{2}} &\quad m < 0 \\
        \end{cases}

    Examples
    ========

    >>> from sympy import Znm, Symbol, simplify
    >>> from sympy.abc import n, m
    >>> theta = Symbol("theta")
    >>> phi = Symbol("phi")
    >>> Znm(n, m, theta, phi)
    Znm(n, m, theta, phi)

    For specific integers n and m we can evaluate the harmonics
    to more useful expressions:

    >>> simplify(Znm(0, 0, theta, phi).expand(func=True))
    1/(2*sqrt(pi))
    >>> simplify(Znm(1, 1, theta, phi).expand(func=True))
    -sqrt(3)*sin(theta)*cos(phi)/(2*sqrt(pi))
    >>> simplify(Znm(2, 1, theta, phi).expand(func=True))
    -sqrt(15)*sin(2*theta)*cos(phi)/(4*sqrt(pi))

    See Also
    ========

    Ynm, Ynm_c

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Spherical_harmonics
    .. [2] https://mathworld.wolfram.com/SphericalHarmonic.html
    .. [3] https://functions.wolfram.com/Polynomials/SphericalHarmonicY/

    """

    @classmethod
    def eval(cls, n, m, theta, phi):
        if m.is_positive:
            zz = (Ynm(n, m, theta, phi) + Ynm_c(n, m, theta, phi)) / sqrt(2)
            return zz
        elif m.is_zero:
            return Ynm(n, m, theta, phi)
        elif m.is_negative:
            zz = (Ynm(n, m, theta, phi) - Ynm_c(n, m, theta, phi)) / (sqrt(2)*I)
            return zz

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\truststore\_api.py ===
import os
import platform
import socket
import ssl
import sys
import typing

import _ssl

from ._ssl_constants import (
    _original_SSLContext,
    _original_super_SSLContext,
    _truststore_SSLContext_dunder_class,
    _truststore_SSLContext_super_class,
)

if platform.system() == "Windows":
    from ._windows import _configure_context, _verify_peercerts_impl
elif platform.system() == "Darwin":
    from ._macos import _configure_context, _verify_peercerts_impl
else:
    from ._openssl import _configure_context, _verify_peercerts_impl

if typing.TYPE_CHECKING:
    from pip._vendor.typing_extensions import Buffer

# From typeshed/stdlib/ssl.pyi
_StrOrBytesPath: typing.TypeAlias = str | bytes | os.PathLike[str] | os.PathLike[bytes]
_PasswordType: typing.TypeAlias = str | bytes | typing.Callable[[], str | bytes]


def inject_into_ssl() -> None:
    """Injects the :class:`truststore.SSLContext` into the ``ssl``
    module by replacing :class:`ssl.SSLContext`.
    """
    setattr(ssl, "SSLContext", SSLContext)
    # urllib3 holds on to its own reference of ssl.SSLContext
    # so we need to replace that reference too.
    try:
        import pip._vendor.urllib3.util.ssl_ as urllib3_ssl

        setattr(urllib3_ssl, "SSLContext", SSLContext)
    except ImportError:
        pass

    # requests starting with 2.32.0 added a preloaded SSL context to improve concurrent performance;
    # this unfortunately leads to a RecursionError, which can be avoided by patching the preloaded SSL context with
    # the truststore patched instance
    # also see https://github.com/psf/requests/pull/6667
    try:
        from pip._vendor.requests import adapters as requests_adapters

        preloaded_context = getattr(requests_adapters, "_preloaded_ssl_context", None)
        if preloaded_context is not None:
            setattr(
                requests_adapters,
                "_preloaded_ssl_context",
                SSLContext(ssl.PROTOCOL_TLS_CLIENT),
            )
    except ImportError:
        pass


def extract_from_ssl() -> None:
    """Restores the :class:`ssl.SSLContext` class to its original state"""
    setattr(ssl, "SSLContext", _original_SSLContext)
    try:
        import pip._vendor.urllib3.util.ssl_ as urllib3_ssl

        urllib3_ssl.SSLContext = _original_SSLContext  # type: ignore[assignment]
    except ImportError:
        pass


class SSLContext(_truststore_SSLContext_super_class):  # type: ignore[misc]
    """SSLContext API that uses system certificates on all platforms"""

    @property  # type: ignore[misc]
    def __class__(self) -> type:
        # Dirty hack to get around isinstance() checks
        # for ssl.SSLContext instances in aiohttp/trustme
        # when using non-CPython implementations.
        return _truststore_SSLContext_dunder_class or SSLContext

    def __init__(self, protocol: int = None) -> None:  # type: ignore[assignment]
        self._ctx = _original_SSLContext(protocol)

        class TruststoreSSLObject(ssl.SSLObject):
            # This object exists because wrap_bio() doesn't
            # immediately do the handshake so we need to do
            # certificate verifications after SSLObject.do_handshake()

            def do_handshake(self) -> None:
                ret = super().do_handshake()
                _verify_peercerts(self, server_hostname=self.server_hostname)
                return ret

        self._ctx.sslobject_class = TruststoreSSLObject

    def wrap_socket(
        self,
        sock: socket.socket,
        server_side: bool = False,
        do_handshake_on_connect: bool = True,
        suppress_ragged_eofs: bool = True,
        server_hostname: str | None = None,
        session: ssl.SSLSession | None = None,
    ) -> ssl.SSLSocket:
        # Use a context manager here because the
        # inner SSLContext holds on to our state
        # but also does the actual handshake.
        with _configure_context(self._ctx):
            ssl_sock = self._ctx.wrap_socket(
                sock,
                server_side=server_side,
                server_hostname=server_hostname,
                do_handshake_on_connect=do_handshake_on_connect,
                suppress_ragged_eofs=suppress_ragged_eofs,
                session=session,
            )
        try:
            _verify_peercerts(ssl_sock, server_hostname=server_hostname)
        except Exception:
            ssl_sock.close()
            raise
        return ssl_sock

    def wrap_bio(
        self,
        incoming: ssl.MemoryBIO,
        outgoing: ssl.MemoryBIO,
        server_side: bool = False,
        server_hostname: str | None = None,
        session: ssl.SSLSession | None = None,
    ) -> ssl.SSLObject:
        with _configure_context(self._ctx):
            ssl_obj = self._ctx.wrap_bio(
                incoming,
                outgoing,
                server_hostname=server_hostname,
                server_side=server_side,
                session=session,
            )
        return ssl_obj

    def load_verify_locations(
        self,
        cafile: str | bytes | os.PathLike[str] | os.PathLike[bytes] | None = None,
        capath: str | bytes | os.PathLike[str] | os.PathLike[bytes] | None = None,
        cadata: typing.Union[str, "Buffer", None] = None,
    ) -> None:
        return self._ctx.load_verify_locations(
            cafile=cafile, capath=capath, cadata=cadata
        )

    def load_cert_chain(
        self,
        certfile: _StrOrBytesPath,
        keyfile: _StrOrBytesPath | None = None,
        password: _PasswordType | None = None,
    ) -> None:
        return self._ctx.load_cert_chain(
            certfile=certfile, keyfile=keyfile, password=password
        )

    def load_default_certs(
        self, purpose: ssl.Purpose = ssl.Purpose.SERVER_AUTH
    ) -> None:
        return self._ctx.load_default_certs(purpose)

    def set_alpn_protocols(self, alpn_protocols: typing.Iterable[str]) -> None:
        return self._ctx.set_alpn_protocols(alpn_protocols)

    def set_npn_protocols(self, npn_protocols: typing.Iterable[str]) -> None:
        return self._ctx.set_npn_protocols(npn_protocols)

    def set_ciphers(self, __cipherlist: str) -> None:
        return self._ctx.set_ciphers(__cipherlist)

    def get_ciphers(self) -> typing.Any:
        return self._ctx.get_ciphers()

    def session_stats(self) -> dict[str, int]:
        return self._ctx.session_stats()

    def cert_store_stats(self) -> dict[str, int]:
        raise NotImplementedError()

    def set_default_verify_paths(self) -> None:
        self._ctx.set_default_verify_paths()

    @typing.overload
    def get_ca_certs(
        self, binary_form: typing.Literal[False] = ...
    ) -> list[typing.Any]: ...

    @typing.overload
    def get_ca_certs(self, binary_form: typing.Literal[True] = ...) -> list[bytes]: ...

    @typing.overload
    def get_ca_certs(self, binary_form: bool = ...) -> typing.Any: ...

    def get_ca_certs(self, binary_form: bool = False) -> list[typing.Any] | list[bytes]:
        raise NotImplementedError()

    @property
    def check_hostname(self) -> bool:
        return self._ctx.check_hostname

    @check_hostname.setter
    def check_hostname(self, value: bool) -> None:
        self._ctx.check_hostname = value

    @property
    def hostname_checks_common_name(self) -> bool:
        return self._ctx.hostname_checks_common_name

    @hostname_checks_common_name.setter
    def hostname_checks_common_name(self, value: bool) -> None:
        self._ctx.hostname_checks_common_name = value

    @property
    def keylog_filename(self) -> str:
        return self._ctx.keylog_filename

    @keylog_filename.setter
    def keylog_filename(self, value: str) -> None:
        self._ctx.keylog_filename = value

    @property
    def maximum_version(self) -> ssl.TLSVersion:
        return self._ctx.maximum_version

    @maximum_version.setter
    def maximum_version(self, value: ssl.TLSVersion) -> None:
        _original_super_SSLContext.maximum_version.__set__(  # type: ignore[attr-defined]
            self._ctx, value
        )

    @property
    def minimum_version(self) -> ssl.TLSVersion:
        return self._ctx.minimum_version

    @minimum_version.setter
    def minimum_version(self, value: ssl.TLSVersion) -> None:
        _original_super_SSLContext.minimum_version.__set__(  # type: ignore[attr-defined]
            self._ctx, value
        )

    @property
    def options(self) -> ssl.Options:
        return self._ctx.options

    @options.setter
    def options(self, value: ssl.Options) -> None:
        _original_super_SSLContext.options.__set__(  # type: ignore[attr-defined]
            self._ctx, value
        )

    @property
    def post_handshake_auth(self) -> bool:
        return self._ctx.post_handshake_auth

    @post_handshake_auth.setter
    def post_handshake_auth(self, value: bool) -> None:
        self._ctx.post_handshake_auth = value

    @property
    def protocol(self) -> ssl._SSLMethod:
        return self._ctx.protocol

    @property
    def security_level(self) -> int:
        return self._ctx.security_level

    @property
    def verify_flags(self) -> ssl.VerifyFlags:
        return self._ctx.verify_flags

    @verify_flags.setter
    def verify_flags(self, value: ssl.VerifyFlags) -> None:
        _original_super_SSLContext.verify_flags.__set__(  # type: ignore[attr-defined]
            self._ctx, value
        )

    @property
    def verify_mode(self) -> ssl.VerifyMode:
        return self._ctx.verify_mode

    @verify_mode.setter
    def verify_mode(self, value: ssl.VerifyMode) -> None:
        _original_super_SSLContext.verify_mode.__set__(  # type: ignore[attr-defined]
            self._ctx, value
        )


# Python 3.13+ makes get_unverified_chain() a public API that only returns DER
# encoded certificates. We detect whether we need to call public_bytes() for 3.10->3.12
# Pre-3.13 returned None instead of an empty list from get_unverified_chain()
if sys.version_info >= (3, 13):

    def _get_unverified_chain_bytes(sslobj: ssl.SSLObject) -> list[bytes]:
        unverified_chain = sslobj.get_unverified_chain() or ()  # type: ignore[attr-defined]
        return [
            cert if isinstance(cert, bytes) else cert.public_bytes(_ssl.ENCODING_DER)
            for cert in unverified_chain
        ]

else:

    def _get_unverified_chain_bytes(sslobj: ssl.SSLObject) -> list[bytes]:
        unverified_chain = sslobj.get_unverified_chain() or ()  # type: ignore[attr-defined]
        return [cert.public_bytes(_ssl.ENCODING_DER) for cert in unverified_chain]


def _verify_peercerts(
    sock_or_sslobj: ssl.SSLSocket | ssl.SSLObject, server_hostname: str | None
) -> None:
    """
    Verifies the peer certificates from an SSLSocket or SSLObject
    against the certificates in the OS trust store.
    """
    sslobj: ssl.SSLObject = sock_or_sslobj  # type: ignore[assignment]
    try:
        while not hasattr(sslobj, "get_unverified_chain"):
            sslobj = sslobj._sslobj  # type: ignore[attr-defined]
    except AttributeError:
        pass

    cert_bytes = _get_unverified_chain_bytes(sslobj)
    _verify_peercerts_impl(
        sock_or_sslobj.context, cert_bytes, server_hostname=server_hostname
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\cell\cell.py ===
# Copyright (c) 2010-2024 openpyxl

"""Manage individual cells in a spreadsheet.

The Cell class is required to know its value and type, display options,
and any other features of an Excel cell.  Utilities for referencing
cells using Excel's 'A1' column/row nomenclature are also provided.

"""

__docformat__ = "restructuredtext en"

# Python stdlib imports
from copy import copy
import datetime
import re


from openpyxl.compat import (
    NUMERIC_TYPES,
)

from openpyxl.utils.exceptions import IllegalCharacterError

from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers, is_date_format
from openpyxl.styles.styleable import StyleableObject
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.formula import DataTableFormula, ArrayFormula
from openpyxl.cell.rich_text import CellRichText

# constants

TIME_TYPES = (datetime.datetime, datetime.date, datetime.time, datetime.timedelta)
TIME_FORMATS = {
    datetime.datetime:numbers.FORMAT_DATE_DATETIME,
    datetime.date:numbers.FORMAT_DATE_YYYYMMDD2,
    datetime.time:numbers.FORMAT_DATE_TIME6,
    datetime.timedelta:numbers.FORMAT_DATE_TIMEDELTA,
                }

STRING_TYPES = (str, bytes, CellRichText)
KNOWN_TYPES = NUMERIC_TYPES + TIME_TYPES + STRING_TYPES + (bool, type(None))

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
ERROR_CODES = ('#NULL!', '#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#NUM!',
               '#N/A')

TYPE_STRING = 's'
TYPE_FORMULA = 'f'
TYPE_NUMERIC = 'n'
TYPE_BOOL = 'b'
TYPE_NULL = 'n'
TYPE_INLINE = 'inlineStr'
TYPE_ERROR = 'e'
TYPE_FORMULA_CACHE_STRING = 'str'

VALID_TYPES = (TYPE_STRING, TYPE_FORMULA, TYPE_NUMERIC, TYPE_BOOL,
               TYPE_NULL, TYPE_INLINE, TYPE_ERROR, TYPE_FORMULA_CACHE_STRING)


_TYPES = {int:'n', float:'n', str:'s', bool:'b'}


def get_type(t, value):
    if isinstance(value, NUMERIC_TYPES):
        dt = 'n'
    elif isinstance(value, STRING_TYPES):
        dt = 's'
    elif isinstance(value, TIME_TYPES):
        dt = 'd'
    elif isinstance(value, (DataTableFormula, ArrayFormula)):
        dt = 'f'
    else:
        return
    _TYPES[t] = dt
    return dt


def get_time_format(t):
    value = TIME_FORMATS.get(t)
    if value:
        return value
    for base in t.mro()[1:]:
        value = TIME_FORMATS.get(base)
        if value:
            TIME_FORMATS[t] = value
            return value
    raise ValueError("Could not get time format for {0!r}".format(value))


class Cell(StyleableObject):
    """Describes cell associated properties.

    Properties of interest include style, type, value, and address.

    """
    __slots__ = (
        'row',
        'column',
        '_value',
        'data_type',
        'parent',
        '_hyperlink',
        '_comment',
                 )

    def __init__(self, worksheet, row=None, column=None, value=None, style_array=None):
        super().__init__(worksheet, style_array)
        self.row = row
        """Row number of this cell (1-based)"""
        self.column = column
        """Column number of this cell (1-based)"""
        # _value is the stored value, while value is the displayed value
        self._value = None
        self._hyperlink = None
        self.data_type = 'n'
        if value is not None:
            self.value = value
        self._comment = None


    @property
    def coordinate(self):
        """This cell's coordinate (ex. 'A5')"""
        col = get_column_letter(self.column)
        return f"{col}{self.row}"


    @property
    def col_idx(self):
        """The numerical index of the column"""
        return self.column


    @property
    def column_letter(self):
        return get_column_letter(self.column)


    @property
    def encoding(self):
        return self.parent.encoding

    @property
    def base_date(self):
        return self.parent.parent.epoch


    def __repr__(self):
        return "<Cell {0!r}.{1}>".format(self.parent.title, self.coordinate)

    def check_string(self, value):
        """Check string coding, length, and line break character"""
        if value is None:
            return
        # convert to str string
        if not isinstance(value, str):
            value = str(value, self.encoding)
        value = str(value)
        # string must never be longer than 32,767 characters
        # truncate if necessary
        value = value[:32767]
        if next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
            raise IllegalCharacterError(f"{value} cannot be used in worksheets.")
        return value

    def check_error(self, value):
        """Tries to convert Error" else N/A"""
        try:
            return str(value)
        except UnicodeDecodeError:
            return u'#N/A'


    def _bind_value(self, value):
        """Given a value, infer the correct data type"""

        self.data_type = "n"
        t = type(value)
        try:
            dt = _TYPES[t]
        except KeyError:
            dt = get_type(t, value)

        if dt is None and value is not None:
            raise ValueError("Cannot convert {0!r} to Excel".format(value))

        if dt:
            self.data_type = dt

        if dt == 'd':
            if not is_date_format(self.number_format):
                self.number_format = get_time_format(t)

        elif dt == "s" and not isinstance(value, CellRichText):
            value = self.check_string(value)
            if len(value) > 1 and value.startswith("="):
                self.data_type = 'f'
            elif value in ERROR_CODES:
                self.data_type = 'e'

        self._value = value


    @property
    def value(self):
        """Get or set the value held in the cell.

        :type: depends on the value (string, float, int or
            :class:`datetime.datetime`)
        """
        return self._value

    @value.setter
    def value(self, value):
        """Set the value and infer type and display options."""
        self._bind_value(value)

    @property
    def internal_value(self):
        """Always returns the value for excel."""
        return self._value

    @property
    def hyperlink(self):
        """Return the hyperlink target or an empty string"""
        return self._hyperlink


    @hyperlink.setter
    def hyperlink(self, val):
        """Set value and display for hyperlinks in a cell.
        Automatically sets the `value` of the cell with link text,
        but you can modify it afterwards by setting the `value`
        property, and the hyperlink will remain.
        Hyperlink is removed if set to ``None``."""
        if val is None:
            self._hyperlink = None
        else:
            if not isinstance(val, Hyperlink):
                val = Hyperlink(ref="", target=val)
            val.ref = self.coordinate
            self._hyperlink = val
            if self._value is None:
                self.value = val.target or val.location


    @property
    def is_date(self):
        """True if the value is formatted as a date

        :type: bool
        """
        return self.data_type == 'd' or (
            self.data_type == 'n' and is_date_format(self.number_format)
            )


    def offset(self, row=0, column=0):
        """Returns a cell location relative to this cell.

        :param row: number of rows to offset
        :type row: int

        :param column: number of columns to offset
        :type column: int

        :rtype: :class:`openpyxl.cell.Cell`
        """
        offset_column = self.col_idx + column
        offset_row = self.row + row
        return self.parent.cell(column=offset_column, row=offset_row)


    @property
    def comment(self):
        """ Returns the comment associated with this cell

            :type: :class:`openpyxl.comments.Comment`
        """
        return self._comment


    @comment.setter
    def comment(self, value):
        """
        Assign a comment to a cell
        """

        if value is not None:
            if value.parent:
                value = copy(value)
            value.bind(self)
        elif value is None and self._comment:
            self._comment.unbind()
        self._comment = value


class MergedCell(StyleableObject):

    """
    Describes the properties of a cell in a merged cell and helps to
    display the borders of the merged cell.

    The value of a MergedCell is always None.
    """

    __slots__ = ('row', 'column')

    _value = None
    data_type = "n"
    comment = None
    hyperlink = None


    def __init__(self, worksheet, row=None, column=None):
        super().__init__(worksheet)
        self.row = row
        self.column = column


    def __repr__(self):
        return "<MergedCell {0!r}.{1}>".format(self.parent.title, self.coordinate)

    coordinate = Cell.coordinate
    _comment = comment
    value = _value


def WriteOnlyCell(ws=None, value=None):
    return Cell(worksheet=ws, column=1, row=1, value=value)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\tools\numeric.py ===
from __future__ import annotations

from typing import (
    TYPE_CHECKING,
    Literal,
)
import warnings

import numpy as np

from pandas._libs import (
    lib,
    missing as libmissing,
)
from pandas.util._exceptions import find_stack_level
from pandas.util._validators import check_dtype_backend

from pandas.core.dtypes.cast import maybe_downcast_numeric
from pandas.core.dtypes.common import (
    ensure_object,
    is_bool_dtype,
    is_decimal,
    is_integer_dtype,
    is_number,
    is_numeric_dtype,
    is_scalar,
    is_string_dtype,
    needs_i8_conversion,
)
from pandas.core.dtypes.dtypes import ArrowDtype
from pandas.core.dtypes.generic import (
    ABCIndex,
    ABCSeries,
)

from pandas.core.arrays import BaseMaskedArray
from pandas.core.arrays.string_ import StringDtype

if TYPE_CHECKING:
    from pandas._typing import (
        DateTimeErrorChoices,
        DtypeBackend,
        npt,
    )


def to_numeric(
    arg,
    errors: DateTimeErrorChoices = "raise",
    downcast: Literal["integer", "signed", "unsigned", "float"] | None = None,
    dtype_backend: DtypeBackend | lib.NoDefault = lib.no_default,
):
    """
    Convert argument to a numeric type.

    The default return dtype is `float64` or `int64`
    depending on the data supplied. Use the `downcast` parameter
    to obtain other dtypes.

    Please note that precision loss may occur if really large numbers
    are passed in. Due to the internal limitations of `ndarray`, if
    numbers smaller than `-9223372036854775808` (np.iinfo(np.int64).min)
    or larger than `18446744073709551615` (np.iinfo(np.uint64).max) are
    passed in, it is very likely they will be converted to float so that
    they can be stored in an `ndarray`. These warnings apply similarly to
    `Series` since it internally leverages `ndarray`.

    Parameters
    ----------
    arg : scalar, list, tuple, 1-d array, or Series
        Argument to be converted.
    errors : {'ignore', 'raise', 'coerce'}, default 'raise'
        - If 'raise', then invalid parsing will raise an exception.
        - If 'coerce', then invalid parsing will be set as NaN.
        - If 'ignore', then invalid parsing will return the input.

        .. versionchanged:: 2.2

        "ignore" is deprecated. Catch exceptions explicitly instead.

    downcast : str, default None
        Can be 'integer', 'signed', 'unsigned', or 'float'.
        If not None, and if the data has been successfully cast to a
        numerical dtype (or if the data was numeric to begin with),
        downcast that resulting data to the smallest numerical dtype
        possible according to the following rules:

        - 'integer' or 'signed': smallest signed int dtype (min.: np.int8)
        - 'unsigned': smallest unsigned int dtype (min.: np.uint8)
        - 'float': smallest float dtype (min.: np.float32)

        As this behaviour is separate from the core conversion to
        numeric values, any errors raised during the downcasting
        will be surfaced regardless of the value of the 'errors' input.

        In addition, downcasting will only occur if the size
        of the resulting data's dtype is strictly larger than
        the dtype it is to be cast to, so if none of the dtypes
        checked satisfy that specification, no downcasting will be
        performed on the data.
    dtype_backend : {'numpy_nullable', 'pyarrow'}, default 'numpy_nullable'
        Back-end data type applied to the resultant :class:`DataFrame`
        (still experimental). Behaviour is as follows:

        * ``"numpy_nullable"``: returns nullable-dtype-backed :class:`DataFrame`
          (default).
        * ``"pyarrow"``: returns pyarrow-backed nullable :class:`ArrowDtype`
          DataFrame.

        .. versionadded:: 2.0

    Returns
    -------
    ret
        Numeric if parsing succeeded.
        Return type depends on input.  Series if Series, otherwise ndarray.

    See Also
    --------
    DataFrame.astype : Cast argument to a specified dtype.
    to_datetime : Convert argument to datetime.
    to_timedelta : Convert argument to timedelta.
    numpy.ndarray.astype : Cast a numpy array to a specified type.
    DataFrame.convert_dtypes : Convert dtypes.

    Examples
    --------
    Take separate series and convert to numeric, coercing when told to

    >>> s = pd.Series(['1.0', '2', -3])
    >>> pd.to_numeric(s)
    0    1.0
    1    2.0
    2   -3.0
    dtype: float64
    >>> pd.to_numeric(s, downcast='float')
    0    1.0
    1    2.0
    2   -3.0
    dtype: float32
    >>> pd.to_numeric(s, downcast='signed')
    0    1
    1    2
    2   -3
    dtype: int8
    >>> s = pd.Series(['apple', '1.0', '2', -3])
    >>> pd.to_numeric(s, errors='coerce')
    0    NaN
    1    1.0
    2    2.0
    3   -3.0
    dtype: float64

    Downcasting of nullable integer and floating dtypes is supported:

    >>> s = pd.Series([1, 2, 3], dtype="Int64")
    >>> pd.to_numeric(s, downcast="integer")
    0    1
    1    2
    2    3
    dtype: Int8
    >>> s = pd.Series([1.0, 2.1, 3.0], dtype="Float64")
    >>> pd.to_numeric(s, downcast="float")
    0    1.0
    1    2.1
    2    3.0
    dtype: Float32
    """
    if downcast not in (None, "integer", "signed", "unsigned", "float"):
        raise ValueError("invalid downcasting method provided")

    if errors not in ("ignore", "raise", "coerce"):
        raise ValueError("invalid error value specified")
    if errors == "ignore":
        # GH#54467
        warnings.warn(
            "errors='ignore' is deprecated and will raise in a future version. "
            "Use to_numeric without passing `errors` and catch exceptions "
            "explicitly instead",
            FutureWarning,
            stacklevel=find_stack_level(),
        )

    check_dtype_backend(dtype_backend)

    is_series = False
    is_index = False
    is_scalars = False

    if isinstance(arg, ABCSeries):
        is_series = True
        values = arg.values
    elif isinstance(arg, ABCIndex):
        is_index = True
        if needs_i8_conversion(arg.dtype):
            values = arg.view("i8")
        else:
            values = arg.values
    elif isinstance(arg, (list, tuple)):
        values = np.array(arg, dtype="O")
    elif is_scalar(arg):
        if is_decimal(arg):
            return float(arg)
        if is_number(arg):
            return arg
        is_scalars = True
        values = np.array([arg], dtype="O")
    elif getattr(arg, "ndim", 1) > 1:
        raise TypeError("arg must be a list, tuple, 1-d array, or Series")
    else:
        values = arg

    orig_values = values

    # GH33013: for IntegerArray & FloatingArray extract non-null values for casting
    # save mask to reconstruct the full array after casting
    mask: npt.NDArray[np.bool_] | None = None
    if isinstance(values, BaseMaskedArray):
        mask = values._mask
        values = values._data[~mask]

    values_dtype = getattr(values, "dtype", None)
    if isinstance(values_dtype, ArrowDtype):
        mask = values.isna()
        values = values.dropna().to_numpy()
    new_mask: np.ndarray | None = None
    if is_numeric_dtype(values_dtype):
        pass
    elif lib.is_np_dtype(values_dtype, "mM"):
        values = values.view(np.int64)
    else:
        values = ensure_object(values)
        coerce_numeric = errors not in ("ignore", "raise")
        try:
            values, new_mask = lib.maybe_convert_numeric(  # type: ignore[call-overload]
                values,
                set(),
                coerce_numeric=coerce_numeric,
                convert_to_masked_nullable=dtype_backend is not lib.no_default
                or isinstance(values_dtype, StringDtype)
                and values_dtype.na_value is libmissing.NA,
            )
        except (ValueError, TypeError):
            if errors == "raise":
                raise
            values = orig_values

    if new_mask is not None:
        # Remove unnecessary values, is expected later anyway and enables
        # downcasting
        values = values[~new_mask]
    elif (
        dtype_backend is not lib.no_default
        and new_mask is None
        or isinstance(values_dtype, StringDtype)
        and values_dtype.na_value is libmissing.NA
    ):
        new_mask = np.zeros(values.shape, dtype=np.bool_)

    # attempt downcast only if the data has been successfully converted
    # to a numerical dtype and if a downcast method has been specified
    if downcast is not None and is_numeric_dtype(values.dtype):
        typecodes: str | None = None

        if downcast in ("integer", "signed"):
            typecodes = np.typecodes["Integer"]
        elif downcast == "unsigned" and (not len(values) or np.min(values) >= 0):
            typecodes = np.typecodes["UnsignedInteger"]
        elif downcast == "float":
            typecodes = np.typecodes["Float"]

            # pandas support goes only to np.float32,
            # as float dtypes smaller than that are
            # extremely rare and not well supported
            float_32_char = np.dtype(np.float32).char
            float_32_ind = typecodes.index(float_32_char)
            typecodes = typecodes[float_32_ind:]

        if typecodes is not None:
            # from smallest to largest
            for typecode in typecodes:
                dtype = np.dtype(typecode)
                if dtype.itemsize <= values.dtype.itemsize:
                    values = maybe_downcast_numeric(values, dtype)

                    # successful conversion
                    if values.dtype == dtype:
                        break

    # GH33013: for IntegerArray, BooleanArray & FloatingArray need to reconstruct
    # masked array
    if (mask is not None or new_mask is not None) and not is_string_dtype(values.dtype):
        if mask is None or (new_mask is not None and new_mask.shape == mask.shape):
            # GH 52588
            mask = new_mask
        else:
            mask = mask.copy()
        assert isinstance(mask, np.ndarray)
        data = np.zeros(mask.shape, dtype=values.dtype)
        data[~mask] = values

        from pandas.core.arrays import (
            ArrowExtensionArray,
            BooleanArray,
            FloatingArray,
            IntegerArray,
        )

        klass: type[IntegerArray | BooleanArray | FloatingArray]
        if is_integer_dtype(data.dtype):
            klass = IntegerArray
        elif is_bool_dtype(data.dtype):
            klass = BooleanArray
        else:
            klass = FloatingArray
        values = klass(data, mask)

        if dtype_backend == "pyarrow" or isinstance(values_dtype, ArrowDtype):
            values = ArrowExtensionArray(values.__arrow_array__())

    if is_series:
        return arg._constructor(values, index=arg.index, name=arg.name)
    elif is_index:
        # because we want to coerce to numeric if possible,
        # do not use _shallow_copy
        from pandas import Index

        return Index(values, name=arg.name)
    elif is_scalars:
        return values[0]
    else:
        return values

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\wheel_builder.py ===
"""Orchestrator for building wheels from InstallRequirements."""

import logging
import os.path
import re
import shutil
from typing import Iterable, List, Optional, Tuple

from pip._vendor.packaging.utils import canonicalize_name, canonicalize_version
from pip._vendor.packaging.version import InvalidVersion, Version

from pip._internal.cache import WheelCache
from pip._internal.exceptions import InvalidWheelFilename, UnsupportedWheel
from pip._internal.metadata import FilesystemWheel, get_wheel_distribution
from pip._internal.models.link import Link
from pip._internal.models.wheel import Wheel
from pip._internal.operations.build.wheel import build_wheel_pep517
from pip._internal.operations.build.wheel_editable import build_wheel_editable
from pip._internal.operations.build.wheel_legacy import build_wheel_legacy
from pip._internal.req.req_install import InstallRequirement
from pip._internal.utils.logging import indent_log
from pip._internal.utils.misc import ensure_dir, hash_file
from pip._internal.utils.setuptools_build import make_setuptools_clean_args
from pip._internal.utils.subprocess import call_subprocess
from pip._internal.utils.temp_dir import TempDirectory
from pip._internal.utils.urls import path_to_url
from pip._internal.vcs import vcs

logger = logging.getLogger(__name__)

_egg_info_re = re.compile(r"([a-z0-9_.]+)-([a-z0-9_.!+-]+)", re.IGNORECASE)

BuildResult = Tuple[List[InstallRequirement], List[InstallRequirement]]


def _contains_egg_info(s: str) -> bool:
    """Determine whether the string looks like an egg_info.

    :param s: The string to parse. E.g. foo-2.1
    """
    return bool(_egg_info_re.search(s))


def _should_build(
    req: InstallRequirement,
) -> bool:
    """Return whether an InstallRequirement should be built into a wheel."""
    assert not req.constraint

    if req.is_wheel:
        return False

    assert req.source_dir

    if req.editable:
        # we only build PEP 660 editable requirements
        return req.supports_pyproject_editable

    return True


def should_build_for_install_command(
    req: InstallRequirement,
) -> bool:
    return _should_build(req)


def _should_cache(
    req: InstallRequirement,
) -> Optional[bool]:
    """
    Return whether a built InstallRequirement can be stored in the persistent
    wheel cache, assuming the wheel cache is available, and _should_build()
    has determined a wheel needs to be built.
    """
    if req.editable or not req.source_dir:
        # never cache editable requirements
        return False

    if req.link and req.link.is_vcs:
        # VCS checkout. Do not cache
        # unless it points to an immutable commit hash.
        assert not req.editable
        assert req.source_dir
        vcs_backend = vcs.get_backend_for_scheme(req.link.scheme)
        assert vcs_backend
        if vcs_backend.is_immutable_rev_checkout(req.link.url, req.source_dir):
            return True
        return False

    assert req.link
    base, ext = req.link.splitext()
    if _contains_egg_info(base):
        return True

    # Otherwise, do not cache.
    return False


def _get_cache_dir(
    req: InstallRequirement,
    wheel_cache: WheelCache,
) -> str:
    """Return the persistent or temporary cache directory where the built
    wheel need to be stored.
    """
    cache_available = bool(wheel_cache.cache_dir)
    assert req.link
    if cache_available and _should_cache(req):
        cache_dir = wheel_cache.get_path_for_link(req.link)
    else:
        cache_dir = wheel_cache.get_ephem_path_for_link(req.link)
    return cache_dir


def _verify_one(req: InstallRequirement, wheel_path: str) -> None:
    canonical_name = canonicalize_name(req.name or "")
    w = Wheel(os.path.basename(wheel_path))
    if canonicalize_name(w.name) != canonical_name:
        raise InvalidWheelFilename(
            f"Wheel has unexpected file name: expected {canonical_name!r}, "
            f"got {w.name!r}",
        )
    dist = get_wheel_distribution(FilesystemWheel(wheel_path), canonical_name)
    dist_verstr = str(dist.version)
    if canonicalize_version(dist_verstr) != canonicalize_version(w.version):
        raise InvalidWheelFilename(
            f"Wheel has unexpected file name: expected {dist_verstr!r}, "
            f"got {w.version!r}",
        )
    metadata_version_value = dist.metadata_version
    if metadata_version_value is None:
        raise UnsupportedWheel("Missing Metadata-Version")
    try:
        metadata_version = Version(metadata_version_value)
    except InvalidVersion:
        msg = f"Invalid Metadata-Version: {metadata_version_value}"
        raise UnsupportedWheel(msg)
    if metadata_version >= Version("1.2") and not isinstance(dist.version, Version):
        raise UnsupportedWheel(
            f"Metadata 1.2 mandates PEP 440 version, but {dist_verstr!r} is not"
        )


def _build_one(
    req: InstallRequirement,
    output_dir: str,
    verify: bool,
    build_options: List[str],
    global_options: List[str],
    editable: bool,
) -> Optional[str]:
    """Build one wheel.

    :return: The filename of the built wheel, or None if the build failed.
    """
    artifact = "editable" if editable else "wheel"
    try:
        ensure_dir(output_dir)
    except OSError as e:
        logger.warning(
            "Building %s for %s failed: %s",
            artifact,
            req.name,
            e,
        )
        return None

    # Install build deps into temporary directory (PEP 518)
    with req.build_env:
        wheel_path = _build_one_inside_env(
            req, output_dir, build_options, global_options, editable
        )
    if wheel_path and verify:
        try:
            _verify_one(req, wheel_path)
        except (InvalidWheelFilename, UnsupportedWheel) as e:
            logger.warning("Built %s for %s is invalid: %s", artifact, req.name, e)
            return None
    return wheel_path


def _build_one_inside_env(
    req: InstallRequirement,
    output_dir: str,
    build_options: List[str],
    global_options: List[str],
    editable: bool,
) -> Optional[str]:
    with TempDirectory(kind="wheel") as temp_dir:
        assert req.name
        if req.use_pep517:
            assert req.metadata_directory
            assert req.pep517_backend
            if global_options:
                logger.warning(
                    "Ignoring --global-option when building %s using PEP 517", req.name
                )
            if build_options:
                logger.warning(
                    "Ignoring --build-option when building %s using PEP 517", req.name
                )
            if editable:
                wheel_path = build_wheel_editable(
                    name=req.name,
                    backend=req.pep517_backend,
                    metadata_directory=req.metadata_directory,
                    tempd=temp_dir.path,
                )
            else:
                wheel_path = build_wheel_pep517(
                    name=req.name,
                    backend=req.pep517_backend,
                    metadata_directory=req.metadata_directory,
                    tempd=temp_dir.path,
                )
        else:
            wheel_path = build_wheel_legacy(
                name=req.name,
                setup_py_path=req.setup_py_path,
                source_dir=req.unpacked_source_directory,
                global_options=global_options,
                build_options=build_options,
                tempd=temp_dir.path,
            )

        if wheel_path is not None:
            wheel_name = os.path.basename(wheel_path)
            dest_path = os.path.join(output_dir, wheel_name)
            try:
                wheel_hash, length = hash_file(wheel_path)
                shutil.move(wheel_path, dest_path)
                logger.info(
                    "Created wheel for %s: filename=%s size=%d sha256=%s",
                    req.name,
                    wheel_name,
                    length,
                    wheel_hash.hexdigest(),
                )
                logger.info("Stored in directory: %s", output_dir)
                return dest_path
            except Exception as e:
                logger.warning(
                    "Building wheel for %s failed: %s",
                    req.name,
                    e,
                )
        # Ignore return, we can't do anything else useful.
        if not req.use_pep517:
            _clean_one_legacy(req, global_options)
        return None


def _clean_one_legacy(req: InstallRequirement, global_options: List[str]) -> bool:
    clean_args = make_setuptools_clean_args(
        req.setup_py_path,
        global_options=global_options,
    )

    logger.info("Running setup.py clean for %s", req.name)
    try:
        call_subprocess(
            clean_args, command_desc="python setup.py clean", cwd=req.source_dir
        )
        return True
    except Exception:
        logger.error("Failed cleaning build dir for %s", req.name)
        return False


def build(
    requirements: Iterable[InstallRequirement],
    wheel_cache: WheelCache,
    verify: bool,
    build_options: List[str],
    global_options: List[str],
) -> BuildResult:
    """Build wheels.

    :return: The list of InstallRequirement that succeeded to build and
        the list of InstallRequirement that failed to build.
    """
    if not requirements:
        return [], []

    # Build the wheels.
    logger.info(
        "Building wheels for collected packages: %s",
        ", ".join(req.name for req in requirements),  # type: ignore
    )

    with indent_log():
        build_successes, build_failures = [], []
        for req in requirements:
            assert req.name
            cache_dir = _get_cache_dir(req, wheel_cache)
            wheel_file = _build_one(
                req,
                cache_dir,
                verify,
                build_options,
                global_options,
                req.editable and req.permit_editable_wheels,
            )
            if wheel_file:
                # Record the download origin in the cache
                if req.download_info is not None:
                    # download_info is guaranteed to be set because when we build an
                    # InstallRequirement it has been through the preparer before, but
                    # let's be cautious.
                    wheel_cache.record_download_origin(cache_dir, req.download_info)
                # Update the link for this.
                req.link = Link(path_to_url(wheel_file))
                req.local_file_path = req.link.file_path
                assert req.link.is_wheel
                build_successes.append(req)
            else:
                build_failures.append(req)

    # notify success/failure
    if build_successes:
        logger.info(
            "Successfully built %s",
            " ".join([req.name for req in build_successes]),  # type: ignore
        )
    if build_failures:
        logger.info(
            "Failed to build %s",
            " ".join([req.name for req in build_failures]),  # type: ignore
        )
    # Return a list of requirements that failed to build
    return build_successes, build_failures

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\support\color.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.
from __future__ import annotations

import sys
from typing import TYPE_CHECKING, Any, Sequence

if sys.version_info >= (3, 9):
    from re import Match
else:
    from typing import Match

if TYPE_CHECKING:
    from typing import SupportsFloat, SupportsIndex, SupportsInt, Union

    ParseableFloat = Union[SupportsFloat, SupportsIndex, str, bytes, bytearray]
    ParseableInt = Union[SupportsInt, SupportsIndex, str, bytes]
else:
    ParseableFloat = Any
    ParseableInt = Any

RGB_PATTERN = r"^\s*rgb\(\s*(\d{1,3})\s*,\s*(\d{1,3})\s*,\s*(\d{1,3})\s*\)\s*$"
RGB_PCT_PATTERN = (
    r"^\s*rgb\(\s*(\d{1,3}|\d{1,2}\.\d+)%\s*,\s*(\d{1,3}|\d{1,2}\.\d+)%\s*,\s*(\d{1,3}|\d{1,2}\.\d+)%\s*\)\s*$"
)
RGBA_PATTERN = r"^\s*rgba\(\s*(\d{1,3})\s*,\s*(\d{1,3})\s*,\s*(\d{1,3})\s*,\s*(0|1|0\.\d+)\s*\)\s*$"
RGBA_PCT_PATTERN = (
    r"^\s*rgba\(\s*(\d{1,3}|\d{1,2}\.\d+)%\s*,\s*(\d{1,3}|\d{1,2}\.\d+)%\s*,"
    + r"\s*(\d{1,3}|\d{1,2}\.\d+)%\s*,\s*(0|1|0\.\d+)\s*\)\s*$"
)
HEX_PATTERN = r"#([A-Fa-f0-9]{2})([A-Fa-f0-9]{2})([A-Fa-f0-9]{2})"
HEX3_PATTERN = r"#([A-Fa-f0-9])([A-Fa-f0-9])([A-Fa-f0-9])"
HSL_PATTERN = r"^\s*hsl\(\s*(\d{1,3})\s*,\s*(\d{1,3})%\s*,\s*(\d{1,3})%\s*\)\s*$"
HSLA_PATTERN = r"^\s*hsla\(\s*(\d{1,3})\s*,\s*(\d{1,3})%\s*,\s*(\d{1,3})%\s*,\s*(0|1|0\.\d+)\s*\)\s*$"


class Color:
    """Color conversion support class.

    Example:

    ::

        from selenium.webdriver.support.color import Color

        print(Color.from_string("#00ff33").rgba)
        print(Color.from_string("rgb(1, 255, 3)").hex)
        print(Color.from_string("blue").rgba)
    """

    @classmethod
    def from_string(cls, str_: str) -> Color:
        import re

        class Matcher:
            match_obj: Match[str] | None

            def __init__(self) -> None:
                self.match_obj = None

            def match(self, pattern: str, str_: str) -> Match[str] | None:
                self.match_obj = re.match(pattern, str_)
                return self.match_obj

            @property
            def groups(self) -> Sequence[str]:
                return () if not self.match_obj else self.match_obj.groups()

        m = Matcher()

        if m.match(RGB_PATTERN, str_):
            return cls(*m.groups)
        if m.match(RGB_PCT_PATTERN, str_):
            rgb = tuple(float(each) / 100 * 255 for each in m.groups)
            return cls(*rgb)
        if m.match(RGBA_PATTERN, str_):
            return cls(*m.groups)
        if m.match(RGBA_PCT_PATTERN, str_):
            rgba = tuple([float(each) / 100 * 255 for each in m.groups[:3]] + [m.groups[3]])
            return cls(*rgba)
        if m.match(HEX_PATTERN, str_):
            rgb = tuple(int(each, 16) for each in m.groups)
            return cls(*rgb)
        if m.match(HEX3_PATTERN, str_):
            rgb = tuple(int(each * 2, 16) for each in m.groups)
            return cls(*rgb)
        if m.match(HSL_PATTERN, str_) or m.match(HSLA_PATTERN, str_):
            return cls._from_hsl(*m.groups)
        if str_.upper() in Colors:
            return Colors[str_.upper()]
        raise ValueError("Could not convert %s into color" % str_)

    @classmethod
    def _from_hsl(cls, h: ParseableFloat, s: ParseableFloat, light: ParseableFloat, a: ParseableFloat = 1) -> Color:
        h = float(h) / 360
        s = float(s) / 100
        _l = float(light) / 100

        if s == 0:
            r = _l
            g = r
            b = r
        else:
            luminocity2 = _l * (1 + s) if _l < 0.5 else _l + s - _l * s
            luminocity1 = 2 * _l - luminocity2

            def hue_to_rgb(lum1: float, lum2: float, hue: float) -> float:
                if hue < 0.0:
                    hue += 1
                if hue > 1.0:
                    hue -= 1

                if hue < 1.0 / 6.0:
                    return lum1 + (lum2 - lum1) * 6.0 * hue
                if hue < 1.0 / 2.0:
                    return lum2
                if hue < 2.0 / 3.0:
                    return lum1 + (lum2 - lum1) * ((2.0 / 3.0) - hue) * 6.0
                return lum1

            r = hue_to_rgb(luminocity1, luminocity2, h + 1.0 / 3.0)
            g = hue_to_rgb(luminocity1, luminocity2, h)
            b = hue_to_rgb(luminocity1, luminocity2, h - 1.0 / 3.0)

        return cls(round(r * 255), round(g * 255), round(b * 255), a)

    def __init__(self, red: ParseableInt, green: ParseableInt, blue: ParseableInt, alpha: ParseableFloat = 1) -> None:
        self.red = int(red)
        self.green = int(green)
        self.blue = int(blue)
        self.alpha = "1" if float(alpha) == 1 else str(float(alpha) or 0)

    @property
    def rgb(self) -> str:
        return f"rgb({self.red}, {self.green}, {self.blue})"

    @property
    def rgba(self) -> str:
        return f"rgba({self.red}, {self.green}, {self.blue}, {self.alpha})"

    @property
    def hex(self) -> str:
        return f"#{self.red:02x}{self.green:02x}{self.blue:02x}"

    def __eq__(self, other: object) -> bool:
        if isinstance(other, Color):
            return self.rgba == other.rgba
        return NotImplemented

    def __ne__(self, other: Any) -> bool:
        result = self.__eq__(other)
        if result is NotImplemented:
            return result
        return not result

    def __hash__(self) -> int:
        return hash((self.red, self.green, self.blue, self.alpha))

    def __repr__(self) -> str:
        return f"Color(red={self.red}, green={self.green}, blue={self.blue}, alpha={self.alpha})"

    def __str__(self) -> str:
        return f"Color: {self.rgba}"


# Basic, extended and transparent colour keywords as defined by the W3C HTML4 spec
# See http://www.w3.org/TR/css3-color/#html4
Colors = {
    "TRANSPARENT": Color(0, 0, 0, 0),
    "ALICEBLUE": Color(240, 248, 255),
    "ANTIQUEWHITE": Color(250, 235, 215),
    "AQUA": Color(0, 255, 255),
    "AQUAMARINE": Color(127, 255, 212),
    "AZURE": Color(240, 255, 255),
    "BEIGE": Color(245, 245, 220),
    "BISQUE": Color(255, 228, 196),
    "BLACK": Color(0, 0, 0),
    "BLANCHEDALMOND": Color(255, 235, 205),
    "BLUE": Color(0, 0, 255),
    "BLUEVIOLET": Color(138, 43, 226),
    "BROWN": Color(165, 42, 42),
    "BURLYWOOD": Color(222, 184, 135),
    "CADETBLUE": Color(95, 158, 160),
    "CHARTREUSE": Color(127, 255, 0),
    "CHOCOLATE": Color(210, 105, 30),
    "CORAL": Color(255, 127, 80),
    "CORNFLOWERBLUE": Color(100, 149, 237),
    "CORNSILK": Color(255, 248, 220),
    "CRIMSON": Color(220, 20, 60),
    "CYAN": Color(0, 255, 255),
    "DARKBLUE": Color(0, 0, 139),
    "DARKCYAN": Color(0, 139, 139),
    "DARKGOLDENROD": Color(184, 134, 11),
    "DARKGRAY": Color(169, 169, 169),
    "DARKGREEN": Color(0, 100, 0),
    "DARKGREY": Color(169, 169, 169),
    "DARKKHAKI": Color(189, 183, 107),
    "DARKMAGENTA": Color(139, 0, 139),
    "DARKOLIVEGREEN": Color(85, 107, 47),
    "DARKORANGE": Color(255, 140, 0),
    "DARKORCHID": Color(153, 50, 204),
    "DARKRED": Color(139, 0, 0),
    "DARKSALMON": Color(233, 150, 122),
    "DARKSEAGREEN": Color(143, 188, 143),
    "DARKSLATEBLUE": Color(72, 61, 139),
    "DARKSLATEGRAY": Color(47, 79, 79),
    "DARKSLATEGREY": Color(47, 79, 79),
    "DARKTURQUOISE": Color(0, 206, 209),
    "DARKVIOLET": Color(148, 0, 211),
    "DEEPPINK": Color(255, 20, 147),
    "DEEPSKYBLUE": Color(0, 191, 255),
    "DIMGRAY": Color(105, 105, 105),
    "DIMGREY": Color(105, 105, 105),
    "DODGERBLUE": Color(30, 144, 255),
    "FIREBRICK": Color(178, 34, 34),
    "FLORALWHITE": Color(255, 250, 240),
    "FORESTGREEN": Color(34, 139, 34),
    "FUCHSIA": Color(255, 0, 255),
    "GAINSBORO": Color(220, 220, 220),
    "GHOSTWHITE": Color(248, 248, 255),
    "GOLD": Color(255, 215, 0),
    "GOLDENROD": Color(218, 165, 32),
    "GRAY": Color(128, 128, 128),
    "GREY": Color(128, 128, 128),
    "GREEN": Color(0, 128, 0),
    "GREENYELLOW": Color(173, 255, 47),
    "HONEYDEW": Color(240, 255, 240),
    "HOTPINK": Color(255, 105, 180),
    "INDIANRED": Color(205, 92, 92),
    "INDIGO": Color(75, 0, 130),
    "IVORY": Color(255, 255, 240),
    "KHAKI": Color(240, 230, 140),
    "LAVENDER": Color(230, 230, 250),
    "LAVENDERBLUSH": Color(255, 240, 245),
    "LAWNGREEN": Color(124, 252, 0),
    "LEMONCHIFFON": Color(255, 250, 205),
    "LIGHTBLUE": Color(173, 216, 230),
    "LIGHTCORAL": Color(240, 128, 128),
    "LIGHTCYAN": Color(224, 255, 255),
    "LIGHTGOLDENRODYELLOW": Color(250, 250, 210),
    "LIGHTGRAY": Color(211, 211, 211),
    "LIGHTGREEN": Color(144, 238, 144),
    "LIGHTGREY": Color(211, 211, 211),
    "LIGHTPINK": Color(255, 182, 193),
    "LIGHTSALMON": Color(255, 160, 122),
    "LIGHTSEAGREEN": Color(32, 178, 170),
    "LIGHTSKYBLUE": Color(135, 206, 250),
    "LIGHTSLATEGRAY": Color(119, 136, 153),
    "LIGHTSLATEGREY": Color(119, 136, 153),
    "LIGHTSTEELBLUE": Color(176, 196, 222),
    "LIGHTYELLOW": Color(255, 255, 224),
    "LIME": Color(0, 255, 0),
    "LIMEGREEN": Color(50, 205, 50),
    "LINEN": Color(250, 240, 230),
    "MAGENTA": Color(255, 0, 255),
    "MAROON": Color(128, 0, 0),
    "MEDIUMAQUAMARINE": Color(102, 205, 170),
    "MEDIUMBLUE": Color(0, 0, 205),
    "MEDIUMORCHID": Color(186, 85, 211),
    "MEDIUMPURPLE": Color(147, 112, 219),
    "MEDIUMSEAGREEN": Color(60, 179, 113),
    "MEDIUMSLATEBLUE": Color(123, 104, 238),
    "MEDIUMSPRINGGREEN": Color(0, 250, 154),
    "MEDIUMTURQUOISE": Color(72, 209, 204),
    "MEDIUMVIOLETRED": Color(199, 21, 133),
    "MIDNIGHTBLUE": Color(25, 25, 112),
    "MINTCREAM": Color(245, 255, 250),
    "MISTYROSE": Color(255, 228, 225),
    "MOCCASIN": Color(255, 228, 181),
    "NAVAJOWHITE": Color(255, 222, 173),
    "NAVY": Color(0, 0, 128),
    "OLDLACE": Color(253, 245, 230),
    "OLIVE": Color(128, 128, 0),
    "OLIVEDRAB": Color(107, 142, 35),
    "ORANGE": Color(255, 165, 0),
    "ORANGERED": Color(255, 69, 0),
    "ORCHID": Color(218, 112, 214),
    "PALEGOLDENROD": Color(238, 232, 170),
    "PALEGREEN": Color(152, 251, 152),
    "PALETURQUOISE": Color(175, 238, 238),
    "PALEVIOLETRED": Color(219, 112, 147),
    "PAPAYAWHIP": Color(255, 239, 213),
    "PEACHPUFF": Color(255, 218, 185),
    "PERU": Color(205, 133, 63),
    "PINK": Color(255, 192, 203),
    "PLUM": Color(221, 160, 221),
    "POWDERBLUE": Color(176, 224, 230),
    "PURPLE": Color(128, 0, 128),
    "REBECCAPURPLE": Color(128, 51, 153),
    "RED": Color(255, 0, 0),
    "ROSYBROWN": Color(188, 143, 143),
    "ROYALBLUE": Color(65, 105, 225),
    "SADDLEBROWN": Color(139, 69, 19),
    "SALMON": Color(250, 128, 114),
    "SANDYBROWN": Color(244, 164, 96),
    "SEAGREEN": Color(46, 139, 87),
    "SEASHELL": Color(255, 245, 238),
    "SIENNA": Color(160, 82, 45),
    "SILVER": Color(192, 192, 192),
    "SKYBLUE": Color(135, 206, 235),
    "SLATEBLUE": Color(106, 90, 205),
    "SLATEGRAY": Color(112, 128, 144),
    "SLATEGREY": Color(112, 128, 144),
    "SNOW": Color(255, 250, 250),
    "SPRINGGREEN": Color(0, 255, 127),
    "STEELBLUE": Color(70, 130, 180),
    "TAN": Color(210, 180, 140),
    "TEAL": Color(0, 128, 128),
    "THISTLE": Color(216, 191, 216),
    "TOMATO": Color(255, 99, 71),
    "TURQUOISE": Color(64, 224, 208),
    "VIOLET": Color(238, 130, 238),
    "WHEAT": Color(245, 222, 179),
    "WHITE": Color(255, 255, 255),
    "WHITESMOKE": Color(245, 245, 245),
    "YELLOW": Color(255, 255, 0),
    "YELLOWGREEN": Color(154, 205, 50),
}

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\testing\fixtures\mypy.py ===
# testing/fixtures/mypy.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors

from __future__ import annotations

import inspect
import os
from pathlib import Path
import re
import shutil
import sys
import tempfile

from .base import TestBase
from .. import config
from ..assertions import eq_
from ... import util


@config.add_to_marker.mypy
class MypyTest(TestBase):
    __requires__ = ("no_sqlalchemy2_stubs",)

    @config.fixture(scope="function")
    def per_func_cachedir(self):
        yield from self._cachedir()

    @config.fixture(scope="class")
    def cachedir(self):
        yield from self._cachedir()

    def _cachedir(self):
        # as of mypy 0.971 i think we need to keep mypy_path empty
        mypy_path = ""

        with tempfile.TemporaryDirectory() as cachedir:
            with open(
                Path(cachedir) / "sqla_mypy_config.cfg", "w"
            ) as config_file:
                config_file.write(
                    f"""
                    [mypy]\n
                    plugins = sqlalchemy.ext.mypy.plugin\n
                    show_error_codes = True\n
                    {mypy_path}
                    disable_error_code = no-untyped-call

                    [mypy-sqlalchemy.*]
                    ignore_errors = True

                    """
                )
            with open(
                Path(cachedir) / "plain_mypy_config.cfg", "w"
            ) as config_file:
                config_file.write(
                    f"""
                    [mypy]\n
                    show_error_codes = True\n
                    {mypy_path}
                    disable_error_code = var-annotated,no-untyped-call
                    [mypy-sqlalchemy.*]
                    ignore_errors = True

                    """
                )
            yield cachedir

    @config.fixture()
    def mypy_runner(self, cachedir):
        from mypy import api

        def run(path, use_plugin=False, use_cachedir=None):
            if use_cachedir is None:
                use_cachedir = cachedir
            args = [
                "--strict",
                "--raise-exceptions",
                "--cache-dir",
                use_cachedir,
                "--config-file",
                os.path.join(
                    use_cachedir,
                    (
                        "sqla_mypy_config.cfg"
                        if use_plugin
                        else "plain_mypy_config.cfg"
                    ),
                ),
            ]

            # mypy as of 0.990 is more aggressively blocking messaging
            # for paths that are in sys.path, and as pytest puts currdir,
            # test/ etc in sys.path, just copy the source file to the
            # tempdir we are working in so that we don't have to try to
            # manipulate sys.path and/or guess what mypy is doing
            filename = os.path.basename(path)
            test_program = os.path.join(use_cachedir, filename)
            if path != test_program:
                shutil.copyfile(path, test_program)
            args.append(test_program)

            # I set this locally but for the suite here needs to be
            # disabled
            os.environ.pop("MYPY_FORCE_COLOR", None)

            stdout, stderr, exitcode = api.run(args)
            return stdout, stderr, exitcode

        return run

    @config.fixture
    def mypy_typecheck_file(self, mypy_runner):
        def run(path, use_plugin=False):
            expected_messages = self._collect_messages(path)
            stdout, stderr, exitcode = mypy_runner(path, use_plugin=use_plugin)
            self._check_output(
                path, expected_messages, stdout, stderr, exitcode
            )

        return run

    @staticmethod
    def file_combinations(dirname):
        if os.path.isabs(dirname):
            path = dirname
        else:
            caller_path = inspect.stack()[1].filename
            path = os.path.join(os.path.dirname(caller_path), dirname)
        files = list(Path(path).glob("**/*.py"))

        for extra_dir in config.options.mypy_extra_test_paths:
            if extra_dir and os.path.isdir(extra_dir):
                files.extend((Path(extra_dir) / dirname).glob("**/*.py"))
        return files

    def _collect_messages(self, path):
        from sqlalchemy.ext.mypy.util import mypy_14

        expected_messages = []
        expected_re = re.compile(
            r"\s*# EXPECTED(_MYPY)?(_RE)?(_ROW)?(_TYPE)?: (.+)"
        )
        py_ver_re = re.compile(r"^#\s*PYTHON_VERSION\s?>=\s?(\d+\.\d+)")
        with open(path) as file_:
            current_assert_messages = []
            for num, line in enumerate(file_, 1):
                m = py_ver_re.match(line)
                if m:
                    major, _, minor = m.group(1).partition(".")
                    if sys.version_info < (int(major), int(minor)):
                        config.skip_test(
                            "Requires python >= %s" % (m.group(1))
                        )
                    continue

                m = expected_re.match(line)
                if m:
                    is_mypy = bool(m.group(1))
                    is_re = bool(m.group(2))
                    is_row = bool(m.group(3))
                    is_type = bool(m.group(4))

                    expected_msg = re.sub(r"# noqa[:]? ?.*", "", m.group(5))
                    if is_row:
                        expected_msg = re.sub(
                            r"Row\[([^\]]+)\]",
                            lambda m: f"tuple[{m.group(1)}, fallback=s"
                            f"qlalchemy.engine.row.{m.group(0)}]",
                            expected_msg,
                        )
                        # For some reason it does not use or syntax (|)
                        expected_msg = re.sub(
                            r"Optional\[(.*)\]",
                            lambda m: f"Union[{m.group(1)}, None]",
                            expected_msg,
                        )

                    if is_type:
                        if not is_re:
                            # the goal here is that we can cut-and-paste
                            # from vscode -> pylance into the
                            # EXPECTED_TYPE: line, then the test suite will
                            # validate that line against what mypy produces
                            expected_msg = re.sub(
                                r"([\[\]])",
                                lambda m: rf"\{m.group(0)}",
                                expected_msg,
                            )

                            # note making sure preceding text matches
                            # with a dot, so that an expect for "Select"
                            # does not match "TypedSelect"
                            expected_msg = re.sub(
                                r"([\w_]+)",
                                lambda m: rf"(?:.*\.)?{m.group(1)}\*?",
                                expected_msg,
                            )

                            expected_msg = re.sub(
                                "List", "builtins.list", expected_msg
                            )

                            expected_msg = re.sub(
                                r"\b(int|str|float|bool)\b",
                                lambda m: rf"builtins.{m.group(0)}\*?",
                                expected_msg,
                            )
                            # expected_msg = re.sub(
                            #     r"(Sequence|Tuple|List|Union)",
                            #     lambda m: fr"typing.{m.group(0)}\*?",
                            #     expected_msg,
                            # )

                        is_mypy = is_re = True
                        expected_msg = f'Revealed type is "{expected_msg}"'

                    if mypy_14 and util.py39:
                        # use_lowercase_names, py39 and above
                        # https://github.com/python/mypy/blob/304997bfb85200fb521ac727ee0ce3e6085e5278/mypy/options.py#L363  # noqa: E501

                        # skip first character which could be capitalized
                        # "List item x not found" type of message
                        expected_msg = expected_msg[0] + re.sub(
                            (
                                r"\b(List|Tuple|Dict|Set)\b"
                                if is_type
                                else r"\b(List|Tuple|Dict|Set|Type)\b"
                            ),
                            lambda m: m.group(1).lower(),
                            expected_msg[1:],
                        )

                    if mypy_14 and util.py310:
                        # use_or_syntax, py310 and above
                        # https://github.com/python/mypy/blob/304997bfb85200fb521ac727ee0ce3e6085e5278/mypy/options.py#L368  # noqa: E501
                        expected_msg = re.sub(
                            r"Optional\[(.*?)\]",
                            lambda m: f"{m.group(1)} | None",
                            expected_msg,
                        )
                    current_assert_messages.append(
                        (is_mypy, is_re, expected_msg.strip())
                    )
                elif current_assert_messages:
                    expected_messages.extend(
                        (num, is_mypy, is_re, expected_msg)
                        for (
                            is_mypy,
                            is_re,
                            expected_msg,
                        ) in current_assert_messages
                    )
                    current_assert_messages[:] = []

        return expected_messages

    def _check_output(
        self, path, expected_messages, stdout: str, stderr, exitcode
    ):
        not_located = []
        filename = os.path.basename(path)
        if expected_messages:
            # mypy 0.990 changed how return codes work, so don't assume a
            # 1 or a 0 return code here, could be either depending on if
            # errors were generated or not

            output = []

            raw_lines = stdout.split("\n")
            while raw_lines:
                e = raw_lines.pop(0)
                if re.match(r".+\.py:\d+: error: .*", e):
                    output.append(("error", e))
                elif re.match(
                    r".+\.py:\d+: note: +(?:Possible overload|def ).*", e
                ):
                    while raw_lines:
                        ol = raw_lines.pop(0)
                        if not re.match(r".+\.py:\d+: note: +def .*", ol):
                            raw_lines.insert(0, ol)
                            break
                elif re.match(
                    r".+\.py:\d+: note: .*(?:perhaps|suggestion)", e, re.I
                ):
                    pass
                elif re.match(r".+\.py:\d+: note: .*", e):
                    output.append(("note", e))

            for num, is_mypy, is_re, msg in expected_messages:
                msg = msg.replace("'", '"')
                prefix = "[SQLAlchemy Mypy plugin] " if not is_mypy else ""
                for idx, (typ, errmsg) in enumerate(output):
                    if is_re:
                        if re.match(
                            rf".*{filename}\:{num}\: {typ}\: {prefix}{msg}",
                            errmsg,
                        ):
                            break
                    elif (
                        f"{filename}:{num}: {typ}: {prefix}{msg}"
                        in errmsg.replace("'", '"')
                    ):
                        break
                else:
                    not_located.append(msg)
                    continue
                del output[idx]

            if not_located:
                missing = "\n".join(not_located)
                print("Couldn't locate expected messages:", missing, sep="\n")
                if output:
                    extra = "\n".join(msg for _, msg in output)
                    print("Remaining messages:", extra, sep="\n")
                assert False, "expected messages not found, see stdout"

            if output:
                print(f"{len(output)} messages from mypy were not consumed:")
                print("\n".join(msg for _, msg in output))
                assert False, "errors and/or notes remain, see stdout"

        else:
            if exitcode != 0:
                print(stdout, stderr, sep="\n")

            eq_(exitcode, 0, msg=stdout)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\combinatorics\named_groups.py ===
from sympy.combinatorics.group_constructs import DirectProduct
from sympy.combinatorics.perm_groups import PermutationGroup
from sympy.combinatorics.permutations import Permutation

_af_new = Permutation._af_new


def AbelianGroup(*cyclic_orders):
    """
    Returns the direct product of cyclic groups with the given orders.

    Explanation
    ===========

    According to the structure theorem for finite abelian groups ([1]),
    every finite abelian group can be written as the direct product of
    finitely many cyclic groups.

    Examples
    ========

    >>> from sympy.combinatorics.named_groups import AbelianGroup
    >>> AbelianGroup(3, 4)
    PermutationGroup([
            (6)(0 1 2),
            (3 4 5 6)])
    >>> _.is_group
    True

    See Also
    ========

    DirectProduct

    References
    ==========

    .. [1] https://groupprops.subwiki.org/wiki/Structure_theorem_for_finitely_generated_abelian_groups

    """
    groups = []
    degree = 0
    order = 1
    for size in cyclic_orders:
        degree += size
        order *= size
        groups.append(CyclicGroup(size))
    G = DirectProduct(*groups)
    G._is_abelian = True
    G._degree = degree
    G._order = order

    return G


def AlternatingGroup(n):
    """
    Generates the alternating group on ``n`` elements as a permutation group.

    Explanation
    ===========

    For ``n > 2``, the generators taken are ``(0 1 2), (0 1 2 ... n-1)`` for
    ``n`` odd
    and ``(0 1 2), (1 2 ... n-1)`` for ``n`` even (See [1], p.31, ex.6.9.).
    After the group is generated, some of its basic properties are set.
    The cases ``n = 1, 2`` are handled separately.

    Examples
    ========

    >>> from sympy.combinatorics.named_groups import AlternatingGroup
    >>> G = AlternatingGroup(4)
    >>> G.is_group
    True
    >>> a = list(G.generate_dimino())
    >>> len(a)
    12
    >>> all(perm.is_even for perm in a)
    True

    See Also
    ========

    SymmetricGroup, CyclicGroup, DihedralGroup

    References
    ==========

    .. [1] Armstrong, M. "Groups and Symmetry"

    """
    # small cases are special
    if n in (1, 2):
        return PermutationGroup([Permutation([0])])

    a = list(range(n))
    a[0], a[1], a[2] = a[1], a[2], a[0]
    gen1 = a
    if n % 2:
        a = list(range(1, n))
        a.append(0)
        gen2 = a
    else:
        a = list(range(2, n))
        a.append(1)
        a.insert(0, 0)
        gen2 = a
    gens = [gen1, gen2]
    if gen1 == gen2:
        gens = gens[:1]
    G = PermutationGroup([_af_new(a) for a in gens], dups=False)

    set_alternating_group_properties(G, n, n)
    G._is_alt = True
    return G


def set_alternating_group_properties(G, n, degree):
    """Set known properties of an alternating group. """
    if n < 4:
        G._is_abelian = True
        G._is_nilpotent = True
    else:
        G._is_abelian = False
        G._is_nilpotent = False
    if n < 5:
        G._is_solvable = True
    else:
        G._is_solvable = False
    G._degree = degree
    G._is_transitive = True
    G._is_dihedral = False


def CyclicGroup(n):
    """
    Generates the cyclic group of order ``n`` as a permutation group.

    Explanation
    ===========

    The generator taken is the ``n``-cycle ``(0 1 2 ... n-1)``
    (in cycle notation). After the group is generated, some of its basic
    properties are set.

    Examples
    ========

    >>> from sympy.combinatorics.named_groups import CyclicGroup
    >>> G = CyclicGroup(6)
    >>> G.is_group
    True
    >>> G.order()
    6
    >>> list(G.generate_schreier_sims(af=True))
    [[0, 1, 2, 3, 4, 5], [1, 2, 3, 4, 5, 0], [2, 3, 4, 5, 0, 1],
    [3, 4, 5, 0, 1, 2], [4, 5, 0, 1, 2, 3], [5, 0, 1, 2, 3, 4]]

    See Also
    ========

    SymmetricGroup, DihedralGroup, AlternatingGroup

    """
    a = list(range(1, n))
    a.append(0)
    gen = _af_new(a)
    G = PermutationGroup([gen])

    G._is_abelian = True
    G._is_nilpotent = True
    G._is_solvable = True
    G._degree = n
    G._is_transitive = True
    G._order = n
    G._is_dihedral = (n == 2)
    return G


def DihedralGroup(n):
    r"""
    Generates the dihedral group `D_n` as a permutation group.

    Explanation
    ===========

    The dihedral group `D_n` is the group of symmetries of the regular
    ``n``-gon. The generators taken are the ``n``-cycle ``a = (0 1 2 ... n-1)``
    (a rotation of the ``n``-gon) and ``b = (0 n-1)(1 n-2)...``
    (a reflection of the ``n``-gon) in cycle rotation. It is easy to see that
    these satisfy ``a**n = b**2 = 1`` and ``bab = ~a`` so they indeed generate
    `D_n` (See [1]). After the group is generated, some of its basic properties
    are set.

    Examples
    ========

    >>> from sympy.combinatorics.named_groups import DihedralGroup
    >>> G = DihedralGroup(5)
    >>> G.is_group
    True
    >>> a = list(G.generate_dimino())
    >>> [perm.cyclic_form for perm in a]
    [[], [[0, 1, 2, 3, 4]], [[0, 2, 4, 1, 3]],
    [[0, 3, 1, 4, 2]], [[0, 4, 3, 2, 1]], [[0, 4], [1, 3]],
    [[1, 4], [2, 3]], [[0, 1], [2, 4]], [[0, 2], [3, 4]],
    [[0, 3], [1, 2]]]

    See Also
    ========

    SymmetricGroup, CyclicGroup, AlternatingGroup

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Dihedral_group

    """
    # small cases are special
    if n == 1:
        return PermutationGroup([Permutation([1, 0])])
    if n == 2:
        return PermutationGroup([Permutation([1, 0, 3, 2]),
               Permutation([2, 3, 0, 1]), Permutation([3, 2, 1, 0])])

    a = list(range(1, n))
    a.append(0)
    gen1 = _af_new(a)
    a = list(range(n))
    a.reverse()
    gen2 = _af_new(a)
    G = PermutationGroup([gen1, gen2])
    # if n is a power of 2, group is nilpotent
    if n & (n-1) == 0:
        G._is_nilpotent = True
    else:
        G._is_nilpotent = False
    G._is_dihedral = True
    G._is_abelian = False
    G._is_solvable = True
    G._degree = n
    G._is_transitive = True
    G._order = 2*n
    return G


def SymmetricGroup(n):
    """
    Generates the symmetric group on ``n`` elements as a permutation group.

    Explanation
    ===========

    The generators taken are the ``n``-cycle
    ``(0 1 2 ... n-1)`` and the transposition ``(0 1)`` (in cycle notation).
    (See [1]). After the group is generated, some of its basic properties
    are set.

    Examples
    ========

    >>> from sympy.combinatorics.named_groups import SymmetricGroup
    >>> G = SymmetricGroup(4)
    >>> G.is_group
    True
    >>> G.order()
    24
    >>> list(G.generate_schreier_sims(af=True))
    [[0, 1, 2, 3], [1, 2, 3, 0], [2, 3, 0, 1], [3, 1, 2, 0], [0, 2, 3, 1],
    [1, 3, 0, 2], [2, 0, 1, 3], [3, 2, 0, 1], [0, 3, 1, 2], [1, 0, 2, 3],
    [2, 1, 3, 0], [3, 0, 1, 2], [0, 1, 3, 2], [1, 2, 0, 3], [2, 3, 1, 0],
    [3, 1, 0, 2], [0, 2, 1, 3], [1, 3, 2, 0], [2, 0, 3, 1], [3, 2, 1, 0],
    [0, 3, 2, 1], [1, 0, 3, 2], [2, 1, 0, 3], [3, 0, 2, 1]]

    See Also
    ========

    CyclicGroup, DihedralGroup, AlternatingGroup

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Symmetric_group#Generators_and_relations

    """
    if n == 1:
        G = PermutationGroup([Permutation([0])])
    elif n == 2:
        G = PermutationGroup([Permutation([1, 0])])
    else:
        a = list(range(1, n))
        a.append(0)
        gen1 = _af_new(a)
        a = list(range(n))
        a[0], a[1] = a[1], a[0]
        gen2 = _af_new(a)
        G = PermutationGroup([gen1, gen2])
    set_symmetric_group_properties(G, n, n)
    G._is_sym = True
    return G


def set_symmetric_group_properties(G, n, degree):
    """Set known properties of a symmetric group. """
    if n < 3:
        G._is_abelian = True
        G._is_nilpotent = True
    else:
        G._is_abelian = False
        G._is_nilpotent = False
    if n < 5:
        G._is_solvable = True
    else:
        G._is_solvable = False
    G._degree = degree
    G._is_transitive = True
    G._is_dihedral = (n in [2, 3])  # cf Landau's func and Stirling's approx


def RubikGroup(n):
    """Return a group of Rubik's cube generators

    >>> from sympy.combinatorics.named_groups import RubikGroup
    >>> RubikGroup(2).is_group
    True
    """
    from sympy.combinatorics.generators import rubik
    if n <= 1:
        raise ValueError("Invalid cube. n has to be greater than 1")
    return PermutationGroup(rubik(n))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\printing\jscode.py ===
"""
Javascript code printer

The JavascriptCodePrinter converts single SymPy expressions into single
Javascript expressions, using the functions defined in the Javascript
Math object where possible.

"""

from __future__ import annotations
from typing import Any

from sympy.core import S
from sympy.core.numbers import equal_valued
from sympy.printing.codeprinter import CodePrinter
from sympy.printing.precedence import precedence, PRECEDENCE


# dictionary mapping SymPy function to (argument_conditions, Javascript_function).
# Used in JavascriptCodePrinter._print_Function(self)
known_functions = {
    'Abs': 'Math.abs',
    'acos': 'Math.acos',
    'acosh': 'Math.acosh',
    'asin': 'Math.asin',
    'asinh': 'Math.asinh',
    'atan': 'Math.atan',
    'atan2': 'Math.atan2',
    'atanh': 'Math.atanh',
    'ceiling': 'Math.ceil',
    'cos': 'Math.cos',
    'cosh': 'Math.cosh',
    'exp': 'Math.exp',
    'floor': 'Math.floor',
    'log': 'Math.log',
    'Max': 'Math.max',
    'Min': 'Math.min',
    'sign': 'Math.sign',
    'sin': 'Math.sin',
    'sinh': 'Math.sinh',
    'tan': 'Math.tan',
    'tanh': 'Math.tanh',
}


class JavascriptCodePrinter(CodePrinter):
    """"A Printer to convert Python expressions to strings of JavaScript code
    """
    printmethod = '_javascript'
    language = 'JavaScript'

    _default_settings: dict[str, Any] = dict(CodePrinter._default_settings, **{
        'precision': 17,
        'user_functions': {},
        'contract': True,
    })

    def __init__(self, settings={}):
        CodePrinter.__init__(self, settings)
        self.known_functions = dict(known_functions)
        userfuncs = settings.get('user_functions', {})
        self.known_functions.update(userfuncs)

    def _rate_index_position(self, p):
        return p*5

    def _get_statement(self, codestring):
        return "%s;" % codestring

    def _get_comment(self, text):
        return "// {}".format(text)

    def _declare_number_const(self, name, value):
        return "var {} = {};".format(name, value.evalf(self._settings['precision']))

    def _format_code(self, lines):
        return self.indent_code(lines)

    def _traverse_matrix_indices(self, mat):
        rows, cols = mat.shape
        return ((i, j) for i in range(rows) for j in range(cols))

    def _get_loop_opening_ending(self, indices):
        open_lines = []
        close_lines = []
        loopstart = "for (var %(varble)s=%(start)s; %(varble)s<%(end)s; %(varble)s++){"
        for i in indices:
            # Javascript arrays start at 0 and end at dimension-1
            open_lines.append(loopstart % {
                'varble': self._print(i.label),
                'start': self._print(i.lower),
                'end': self._print(i.upper + 1)})
            close_lines.append("}")
        return open_lines, close_lines

    def _print_Pow(self, expr):
        PREC = precedence(expr)
        if equal_valued(expr.exp, -1):
            return '1/%s' % (self.parenthesize(expr.base, PREC))
        elif equal_valued(expr.exp, 0.5):
            return 'Math.sqrt(%s)' % self._print(expr.base)
        elif expr.exp == S.One/3:
            return 'Math.cbrt(%s)' % self._print(expr.base)
        else:
            return 'Math.pow(%s, %s)' % (self._print(expr.base),
                                 self._print(expr.exp))

    def _print_Rational(self, expr):
        p, q = int(expr.p), int(expr.q)
        return '%d/%d' % (p, q)

    def _print_Mod(self, expr):
        num, den = expr.args
        PREC = precedence(expr)
        snum, sden = [self.parenthesize(arg, PREC) for arg in expr.args]
        # % is remainder (same sign as numerator), not modulo (same sign as
        # denominator), in js. Hence, % only works as modulo if both numbers
        # have the same sign
        if (num.is_nonnegative and den.is_nonnegative or
            num.is_nonpositive and den.is_nonpositive):
            return f"{snum} % {sden}"
        return f"(({snum} % {sden}) + {sden}) % {sden}"

    def _print_Relational(self, expr):
        lhs_code = self._print(expr.lhs)
        rhs_code = self._print(expr.rhs)
        op = expr.rel_op
        return "{} {} {}".format(lhs_code, op, rhs_code)

    def _print_Indexed(self, expr):
        # calculate index for 1d array
        dims = expr.shape
        elem = S.Zero
        offset = S.One
        for i in reversed(range(expr.rank)):
            elem += expr.indices[i]*offset
            offset *= dims[i]
        return "%s[%s]" % (self._print(expr.base.label), self._print(elem))

    def _print_Exp1(self, expr):
        return "Math.E"

    def _print_Pi(self, expr):
        return 'Math.PI'

    def _print_Infinity(self, expr):
        return 'Number.POSITIVE_INFINITY'

    def _print_NegativeInfinity(self, expr):
        return 'Number.NEGATIVE_INFINITY'

    def _print_Piecewise(self, expr):
        from sympy.codegen.ast import Assignment
        if expr.args[-1].cond != True:
            # We need the last conditional to be a True, otherwise the resulting
            # function may not return a result.
            raise ValueError("All Piecewise expressions must contain an "
                             "(expr, True) statement to be used as a default "
                             "condition. Without one, the generated "
                             "expression may not evaluate to anything under "
                             "some condition.")
        lines = []
        if expr.has(Assignment):
            for i, (e, c) in enumerate(expr.args):
                if i == 0:
                    lines.append("if (%s) {" % self._print(c))
                elif i == len(expr.args) - 1 and c == True:
                    lines.append("else {")
                else:
                    lines.append("else if (%s) {" % self._print(c))
                code0 = self._print(e)
                lines.append(code0)
                lines.append("}")
            return "\n".join(lines)
        else:
            # The piecewise was used in an expression, need to do inline
            # operators. This has the downside that inline operators will
            # not work for statements that span multiple lines (Matrix or
            # Indexed expressions).
            ecpairs = ["((%s) ? (\n%s\n)\n" % (self._print(c), self._print(e))
                    for e, c in expr.args[:-1]]
            last_line = ": (\n%s\n)" % self._print(expr.args[-1].expr)
            return ": ".join(ecpairs) + last_line + " ".join([")"*len(ecpairs)])

    def _print_MatrixElement(self, expr):
        return "{}[{}]".format(self.parenthesize(expr.parent,
            PRECEDENCE["Atom"], strict=True),
            expr.j + expr.i*expr.parent.shape[1])

    def indent_code(self, code):
        """Accepts a string of code or a list of code lines"""

        if isinstance(code, str):
            code_lines = self.indent_code(code.splitlines(True))
            return ''.join(code_lines)

        tab = "   "
        inc_token = ('{', '(', '{\n', '(\n')
        dec_token = ('}', ')')

        code = [ line.lstrip(' \t') for line in code ]

        increase = [ int(any(map(line.endswith, inc_token))) for line in code ]
        decrease = [ int(any(map(line.startswith, dec_token)))
                     for line in code ]

        pretty = []
        level = 0
        for n, line in enumerate(code):
            if line in ('', '\n'):
                pretty.append(line)
                continue
            level -= decrease[n]
            pretty.append("%s%s" % (tab*level, line))
            level += increase[n]
        return pretty


def jscode(expr, assign_to=None, **settings):
    """Converts an expr to a string of javascript code

    Parameters
    ==========

    expr : Expr
        A SymPy expression to be converted.
    assign_to : optional
        When given, the argument is used as the name of the variable to which
        the expression is assigned. Can be a string, ``Symbol``,
        ``MatrixSymbol``, or ``Indexed`` type. This is helpful in case of
        line-wrapping, or for expressions that generate multi-line statements.
    precision : integer, optional
        The precision for numbers such as pi [default=15].
    user_functions : dict, optional
        A dictionary where keys are ``FunctionClass`` instances and values are
        their string representations. Alternatively, the dictionary value can
        be a list of tuples i.e. [(argument_test, js_function_string)]. See
        below for examples.
    human : bool, optional
        If True, the result is a single string that may contain some constant
        declarations for the number symbols. If False, the same information is
        returned in a tuple of (symbols_to_declare, not_supported_functions,
        code_text). [default=True].
    contract: bool, optional
        If True, ``Indexed`` instances are assumed to obey tensor contraction
        rules and the corresponding nested loops over indices are generated.
        Setting contract=False will not generate loops, instead the user is
        responsible to provide values for the indices in the code.
        [default=True].

    Examples
    ========

    >>> from sympy import jscode, symbols, Rational, sin, ceiling, Abs
    >>> x, tau = symbols("x, tau")
    >>> jscode((2*tau)**Rational(7, 2))
    '8*Math.sqrt(2)*Math.pow(tau, 7/2)'
    >>> jscode(sin(x), assign_to="s")
    's = Math.sin(x);'

    Custom printing can be defined for certain types by passing a dictionary of
    "type" : "function" to the ``user_functions`` kwarg. Alternatively, the
    dictionary value can be a list of tuples i.e. [(argument_test,
    js_function_string)].

    >>> custom_functions = {
    ...   "ceiling": "CEIL",
    ...   "Abs": [(lambda x: not x.is_integer, "fabs"),
    ...           (lambda x: x.is_integer, "ABS")]
    ... }
    >>> jscode(Abs(x) + ceiling(x), user_functions=custom_functions)
    'fabs(x) + CEIL(x)'

    ``Piecewise`` expressions are converted into conditionals. If an
    ``assign_to`` variable is provided an if statement is created, otherwise
    the ternary operator is used. Note that if the ``Piecewise`` lacks a
    default term, represented by ``(expr, True)`` then an error will be thrown.
    This is to prevent generating an expression that may not evaluate to
    anything.

    >>> from sympy import Piecewise
    >>> expr = Piecewise((x + 1, x > 0), (x, True))
    >>> print(jscode(expr, tau))
    if (x > 0) {
       tau = x + 1;
    }
    else {
       tau = x;
    }

    Support for loops is provided through ``Indexed`` types. With
    ``contract=True`` these expressions will be turned into loops, whereas
    ``contract=False`` will just print the assignment expression that should be
    looped over:

    >>> from sympy import Eq, IndexedBase, Idx
    >>> len_y = 5
    >>> y = IndexedBase('y', shape=(len_y,))
    >>> t = IndexedBase('t', shape=(len_y,))
    >>> Dy = IndexedBase('Dy', shape=(len_y-1,))
    >>> i = Idx('i', len_y-1)
    >>> e=Eq(Dy[i], (y[i+1]-y[i])/(t[i+1]-t[i]))
    >>> jscode(e.rhs, assign_to=e.lhs, contract=False)
    'Dy[i] = (y[i + 1] - y[i])/(t[i + 1] - t[i]);'

    Matrices are also supported, but a ``MatrixSymbol`` of the same dimensions
    must be provided to ``assign_to``. Note that any expression that can be
    generated normally can also exist inside a Matrix:

    >>> from sympy import Matrix, MatrixSymbol
    >>> mat = Matrix([x**2, Piecewise((x + 1, x > 0), (x, True)), sin(x)])
    >>> A = MatrixSymbol('A', 3, 1)
    >>> print(jscode(mat, A))
    A[0] = Math.pow(x, 2);
    if (x > 0) {
       A[1] = x + 1;
    }
    else {
       A[1] = x;
    }
    A[2] = Math.sin(x);
    """

    return JavascriptCodePrinter(settings).doprint(expr, assign_to)


def print_jscode(expr, **settings):
    """Prints the Javascript representation of the given expression.

       See jscode for the meaning of the optional arguments.
    """
    print(jscode(expr, **settings))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\whisper\whisper_chain.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.  See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------

import logging
import os

import onnx
from benchmark_helper import Precision
from convert_generation import (
    get_shared_initializers,
    update_decoder_subgraph_output_cross_attention,
    update_decoder_subgraph_share_buffer_and_use_decoder_masked_mha,
)
from onnx import TensorProto, helper
from transformers import WhisperConfig, WhisperTokenizer

logger = logging.getLogger(__name__)


def verify_inputs(beam_inputs, graph_inputs):
    # Verify that ONNX graph's inputs match beam search op's inputs
    beam_required_inputs = list(filter(lambda beam_input: beam_input, beam_inputs))
    assert len(graph_inputs) == len(beam_required_inputs)
    for graph_input, beam_input in zip(graph_inputs, beam_required_inputs, strict=False):
        # Check if graph_input is in beam_input to handle beam_input names with the "_fp16" suffix
        assert graph_input.name in beam_input


def clean_list(arr, remove_all_strings=True):
    if remove_all_strings:
        # Remove all empty strings in list
        return list(filter(lambda elm: elm != "", arr))

    # Remove empty strings at end of list
    while len(arr) > 0:
        if arr[-1] == "":
            arr.pop()
        else:
            break
    return arr


def chain_model(args):
    # Load encoder/decoder and insert necessary (but unused) graph inputs expected by WhisperBeamSearch op
    encoder_model = onnx.load_model(args.encoder_path, load_external_data=True)
    encoder_model.graph.name = "encoderdecoderinit subgraph"

    decoder_model = onnx.load_model(args.decoder_path, load_external_data=True)
    decoder_model.graph.name = "decoder subgraph"

    config = WhisperConfig.from_pretrained(args.model_name_or_path, cache_dir=args.cache_dir)
    tokenizer = WhisperTokenizer.from_pretrained(args.model_name_or_path, cache_dir=args.cache_dir)

    # Create inputs/outputs for WhisperBeamSearch op
    temperature_name = "temperature_fp16" if args.precision == Precision.FLOAT16 else "temperature"
    beam_inputs = [
        "input_features_fp16" if args.precision == Precision.FLOAT16 else "input_features",
        "max_length",
        "min_length",
        "num_beams",
        "num_return_sequences",
        "length_penalty_fp16" if args.precision == Precision.FLOAT16 else "length_penalty",
        "repetition_penalty_fp16" if args.precision == Precision.FLOAT16 else "repetition_penalty",
        "vocab_mask" if args.use_vocab_mask else "",
        "prefix_vocab_mask" if args.use_prefix_vocab_mask else "",
        "",  # attention mask
        "decoder_input_ids" if args.use_forced_decoder_ids else "",
        "logits_processor" if args.use_logits_processor else "",
        "cross_qk_layer_head" if args.collect_cross_qk else "",
        "extra_decoding_ids" if args.extra_decoding_ids else "",
        temperature_name if args.use_temperature else "",
    ]

    sequence_scores_name = "sequence_scores_fp16" if args.precision == Precision.FLOAT16 else "sequence_scores"
    scores_name = "scores_fp16" if args.precision == Precision.FLOAT16 else "scores"
    beam_outputs = [
        "sequences",
        sequence_scores_name if args.output_sequence_scores else "",
        scores_name if args.output_scores else "",
        "cross_qk" if args.collect_cross_qk else "",
        "no_speech_probs_beam" if args.output_no_speech_probs else "",
    ]

    graph_nodes = []
    if args.precision == Precision.FLOAT16:
        input_features_cast_node = helper.make_node(
            "Cast",
            inputs=["input_features"],
            outputs=["input_features_fp16"],
            name="CastInputFeaturesToFp16",
            to=TensorProto.FLOAT16,
        )
        len_pen_cast_node = helper.make_node(
            "Cast",
            inputs=["length_penalty"],
            outputs=["length_penalty_fp16"],
            name="CastLengthPenaltyToFp16",
            to=TensorProto.FLOAT16,
        )
        rep_pen_cast_node = helper.make_node(
            "Cast",
            inputs=["repetition_penalty"],
            outputs=["repetition_penalty_fp16"],
            name="CastRepetitionPenaltyToFp16",
            to=TensorProto.FLOAT16,
        )
        graph_nodes.extend([input_features_cast_node, len_pen_cast_node, rep_pen_cast_node])

        if args.use_temperature:
            temp_cast_node = helper.make_node(
                "Cast",
                inputs=["temperature"],
                outputs=["temperature_fp16"],
                name="temperature_to_fp16",
                to=TensorProto.FLOAT16,
            )
            graph_nodes.append(temp_cast_node)

        if args.output_sequence_scores:
            output_sequence_scores_cast_node = helper.make_node(
                "Cast",
                inputs=["sequence_scores_fp16"],
                outputs=["sequence_scores"],
                name="CastOutputSequenceScoresToFp32",
                to=TensorProto.FLOAT,
            )
            graph_nodes.append(output_sequence_scores_cast_node)

        if args.output_scores:
            output_scores_cast_node = helper.make_node(
                "Cast",
                inputs=["scores_fp16"],
                outputs=["scores"],
                name="CastScoresToFp32",
                to=TensorProto.FLOAT,
            )
            graph_nodes.append(output_scores_cast_node)

    # Create WhisperBeamSearch op
    beam_search_attrs = [
        helper.make_attribute("eos_token_id", config.eos_token_id),
        helper.make_attribute("pad_token_id", config.pad_token_id),
        helper.make_attribute(
            "decoder_start_token_id", config.decoder_start_token_id
        ),  # same as tokenizer.convert_tokens_to_ids(['<|startoftranscript|>'])[0]
        helper.make_attribute("translate_token_id", tokenizer.convert_tokens_to_ids(["<|translate|>"])[0]),
        helper.make_attribute("transcribe_token_id", tokenizer.convert_tokens_to_ids(["<|transcribe|>"])[0]),
        helper.make_attribute("start_of_lm_token_id", tokenizer.convert_tokens_to_ids(["<|startoflm|>"])[0]),
        (
            helper.make_attribute("no_speech_token_id", tokenizer.convert_tokens_to_ids(["<|nospeech|>"])[0])
            if args.output_no_speech_probs
            else ""
        ),
        helper.make_attribute("no_timestamps_token_id", tokenizer.convert_tokens_to_ids(["<|notimestamps|>"])[0]),
        helper.make_attribute("beginning_timestamp_token_id", tokenizer.convert_tokens_to_ids(["<|0.00|>"])[0]),
        helper.make_attribute("no_repeat_ngram_size", args.no_repeat_ngram_size),
        helper.make_attribute("early_stopping", True),
        helper.make_attribute("model_type", 2),
        helper.make_attribute("decoder_output_cross_qk", 1) if args.collect_cross_qk else "",
    ]
    node = helper.make_node(
        "WhisperBeamSearch",
        inputs=clean_list(beam_inputs, remove_all_strings=False),
        outputs=clean_list(beam_outputs, remove_all_strings=False),
        name="BeamSearch",
        domain="com.microsoft",
    )
    node.attribute.extend(clean_list(beam_search_attrs, remove_all_strings=True))

    # Graph inputs
    input_features = helper.make_tensor_value_info(
        "input_features", TensorProto.FLOAT, ["batch_size", "feature_size", "sequence_length"]
    )
    max_length = helper.make_tensor_value_info("max_length", TensorProto.INT32, [1])
    min_length = helper.make_tensor_value_info("min_length", TensorProto.INT32, [1])
    num_beams = helper.make_tensor_value_info("num_beams", TensorProto.INT32, [1])
    num_return_sequences = helper.make_tensor_value_info("num_return_sequences", TensorProto.INT32, [1])
    length_penalty = helper.make_tensor_value_info("length_penalty", TensorProto.FLOAT, [1])
    repetition_penalty = helper.make_tensor_value_info("repetition_penalty", TensorProto.FLOAT, [1])
    vocab_mask = helper.make_tensor_value_info("vocab_mask", TensorProto.INT32, [config.vocab_size])
    prefix_vocab_mask = helper.make_tensor_value_info(
        "prefix_vocab_mask", TensorProto.INT32, ["batch_size", config.vocab_size]
    )
    decoder_input_ids = helper.make_tensor_value_info(
        "decoder_input_ids", TensorProto.INT32, ["batch_size", "initial_sequence_length"]
    )
    logits_processor = helper.make_tensor_value_info("logits_processor", TensorProto.INT32, [1])
    cross_qk_layer_head = helper.make_tensor_value_info("cross_qk_layer_head", TensorProto.INT32, ["num_layer_head", 2])
    extra_decoding_ids = helper.make_tensor_value_info(
        "extra_decoding_ids", TensorProto.INT32, ["batch_size", "extra_decoding_ids_len"]
    )
    temperature = helper.make_tensor_value_info("temperature", TensorProto.FLOAT, [1])

    graph_inputs = clean_list(
        [
            input_features,
            max_length,
            min_length,
            num_beams,
            num_return_sequences,
            length_penalty,
            repetition_penalty,
            vocab_mask if args.use_vocab_mask else "",
            prefix_vocab_mask if args.use_prefix_vocab_mask else "",
            decoder_input_ids if args.use_forced_decoder_ids else "",
            logits_processor if args.use_logits_processor else "",
            cross_qk_layer_head if args.collect_cross_qk else "",
            extra_decoding_ids if args.extra_decoding_ids else "",
            temperature if args.use_temperature else "",
        ]
    )

    # Graph outputs
    sequences = helper.make_tensor_value_info(
        "sequences", TensorProto.INT32, ["batch_size", "num_return_sequences", "max_length"]
    )
    sequence_scores = helper.make_tensor_value_info("sequence_scores", TensorProto.FLOAT, ["batch_size"])
    scores = helper.make_tensor_value_info("scores", TensorProto.FLOAT, ["batch_size"])
    cross_qk = helper.make_tensor_value_info(
        "cross_qk",
        TensorProto.FLOAT,
        ["batch_size", "num_return_sequences", "num_layer_head_cross_qk", "max_length", "frames"],
    )
    no_speech_probs = helper.make_tensor_value_info("no_speech_probs", TensorProto.FLOAT, ["batch_size"])

    graph_outputs = clean_list(
        [
            sequences,
            sequence_scores if args.output_sequence_scores else "",
            scores if args.output_scores else "",
            cross_qk if args.output_cross_qk or (not args.cross_qk_onnx_model and args.collect_cross_qk) else "",
            no_speech_probs if args.output_no_speech_probs else "",
        ]
    )

    # Replace MultiHeadAttention with DecoderMaskedMultiHeadAttention for CUDA EP inference
    if hasattr(args, "use_gpu") and args.use_gpu:
        if update_decoder_subgraph_share_buffer_and_use_decoder_masked_mha(decoder_model.graph):
            logger.info("Updated whisper decoder subgraph to use DecoderMaskedMultiHeadAttention successfully!")
        else:
            logger.warning("DecoderMaskedMultiHeadAttention could not be applied to whisper decoder subgraph")
        if hasattr(args, "collect_cross_qk") and args.collect_cross_qk:
            update_decoder_subgraph_output_cross_attention(decoder_model.graph)

    # Initializers/opsets
    # Delete shared data between decoder/encoder and move to larger graph initializers
    initializers = get_shared_initializers(encoder_model, decoder_model)
    node.attribute.extend(
        [
            helper.make_attribute("decoder", decoder_model.graph),
            helper.make_attribute("encoder", encoder_model.graph),
        ]
    )

    opset_import = [helper.make_opsetid(domain="com.microsoft", version=1), helper.make_opsetid(domain="", version=17)]

    graph_nodes.append(node)
    if args.output_no_speech_probs:
        prob_cast_node = helper.make_node(
            "Cast",
            inputs=["no_speech_probs_beam"],
            outputs=["no_speech_probs"],
            name="no_speech_probs_cast_to_fp32",
            to=TensorProto.FLOAT,
        )
        graph_nodes.append(prob_cast_node)

    # Make graph with WhisperBeamSearch op
    beam_graph = helper.make_graph(
        graph_nodes,
        name="WhisperBeamSearch Graph",
        inputs=graph_inputs,
        outputs=graph_outputs,
        initializer=initializers,
    )
    beam_graph_input_names = [gi.name for gi in graph_inputs]
    beam_graph_output_names = [go.name for go in graph_outputs]

    if args.cross_qk_onnx_model:
        post_qk_model = onnx.load_model(args.cross_qk_onnx_model, load_external_data=True)
        post_qk_graph = post_qk_model.graph
        beam_graph.initializer.extend(post_qk_graph.initializer)
        beam_graph.node.extend(post_qk_graph.node)
        # If tensor from cross_qk_onnx_model has same name as tensor in beamsearch graph, treat them as same tensor.
        # User should notice this rule when provide cross_qk_onnx_model to append to the beamsearch node.
        for pgi in post_qk_graph.input:
            if (
                (pgi.name not in beam_graph_input_names)
                and (pgi.name not in beam_graph_output_names)
                and (pgi.name != "cross_qk")
            ):
                beam_graph.input.extend([pgi])
        beam_graph.output.extend(post_qk_graph.output)

    # Verify graph's inputs match beam search's inputs
    verify_inputs(beam_inputs, graph_inputs)

    assert decoder_model.ir_version == encoder_model.ir_version
    logger.info(f"Using IR version {decoder_model.ir_version} for chained model")

    # Set IR version of chained model to IR version of subgraphs in order to generate a working E2E model
    beam_model = helper.make_model_gen_version(
        beam_graph,
        producer_name="onnxruntime.transformers",
        opset_imports=opset_import,
        ir_version=decoder_model.ir_version,
    )

    # Save WhisperBeamSearch graph and external data
    if os.path.isfile(args.beam_model_output_dir):
        logger.info(f"Overwriting {args.beam_model_output_dir} and {args.beam_model_output_dir + '.data'}")
        if os.path.exists(args.beam_model_output_dir):
            os.remove(args.beam_model_output_dir)
        if os.path.exists(args.beam_model_output_dir + ".data"):
            os.remove(args.beam_model_output_dir + ".data")

    onnx.save(
        beam_model,
        args.beam_model_output_dir,
        save_as_external_data=args.use_external_data_format,
        all_tensors_to_one_file=True,
        convert_attribute=True,
        location=f"{os.path.basename(args.beam_model_output_dir)}.data",
    )
    try:
        onnx.checker.check_model(args.beam_model_output_dir, full_check=True)
    except Exception as e:
        logger.error(f"An error occurred while running the ONNX checker: {e}", exc_info=True)  # noqa: G201

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\flask_sqlalchemy\model.py ===
from __future__ import annotations

import re
import typing as t

import sqlalchemy as sa
import sqlalchemy.orm as sa_orm

from .query import Query

if t.TYPE_CHECKING:
    from .extension import SQLAlchemy


class _QueryProperty:
    """A class property that creates a query object for a model.

    :meta private:
    """

    def __get__(self, obj: Model | None, cls: type[Model]) -> Query:
        return cls.query_class(
            cls, session=cls.__fsa__.session()  # type: ignore[arg-type]
        )


class Model:
    """The base class of the :attr:`.SQLAlchemy.Model` declarative model class.

    To define models, subclass :attr:`db.Model <.SQLAlchemy.Model>`, not this. To
    customize ``db.Model``, subclass this and pass it as ``model_class`` to
    :class:`.SQLAlchemy`. To customize ``db.Model`` at the metaclass level, pass an
    already created declarative model class as ``model_class``.
    """

    __fsa__: t.ClassVar[SQLAlchemy]
    """Internal reference to the extension object.

    :meta private:
    """

    query_class: t.ClassVar[type[Query]] = Query
    """Query class used by :attr:`query`. Defaults to :attr:`.SQLAlchemy.Query`, which
    defaults to :class:`.Query`.
    """

    query: t.ClassVar[Query] = _QueryProperty()  # type: ignore[assignment]
    """A SQLAlchemy query for a model. Equivalent to ``db.session.query(Model)``. Can be
    customized per-model by overriding :attr:`query_class`.

    .. warning::
        The query interface is considered legacy in SQLAlchemy. Prefer using
        ``session.execute(select())`` instead.
    """

    def __repr__(self) -> str:
        state = sa.inspect(self)
        assert state is not None

        if state.transient:
            pk = f"(transient {id(self)})"
        elif state.pending:
            pk = f"(pending {id(self)})"
        else:
            pk = ", ".join(map(str, state.identity))

        return f"<{type(self).__name__} {pk}>"


class BindMetaMixin(type):
    """Metaclass mixin that sets a model's ``metadata`` based on its ``__bind_key__``.

    If the model sets ``metadata`` or ``__table__`` directly, ``__bind_key__`` is
    ignored. If the ``metadata`` is the same as the parent model, it will not be set
    directly on the child model.
    """

    __fsa__: SQLAlchemy
    metadata: sa.MetaData

    def __init__(
        cls, name: str, bases: tuple[type, ...], d: dict[str, t.Any], **kwargs: t.Any
    ) -> None:
        if not ("metadata" in cls.__dict__ or "__table__" in cls.__dict__):
            bind_key = getattr(cls, "__bind_key__", None)
            parent_metadata = getattr(cls, "metadata", None)
            metadata = cls.__fsa__._make_metadata(bind_key)

            if metadata is not parent_metadata:
                cls.metadata = metadata

        super().__init__(name, bases, d, **kwargs)


class BindMixin:
    """DeclarativeBase mixin to set a model's ``metadata`` based on ``__bind_key__``.

    If no ``__bind_key__`` is specified, the model will use the default metadata
    provided by ``DeclarativeBase`` or ``DeclarativeBaseNoMeta``.
    If the model doesn't set ``metadata`` or ``__table__`` directly
    and does set ``__bind_key__``, the model will use the metadata
    for the specified bind key.
    If the ``metadata`` is the same as the parent model, it will not be set
    directly on the child model.

    .. versionchanged:: 3.1.0
    """

    __fsa__: SQLAlchemy
    metadata: sa.MetaData

    @classmethod
    def __init_subclass__(cls: t.Type[BindMixin], **kwargs: t.Dict[str, t.Any]) -> None:
        if not ("metadata" in cls.__dict__ or "__table__" in cls.__dict__) and hasattr(
            cls, "__bind_key__"
        ):
            bind_key = getattr(cls, "__bind_key__", None)
            parent_metadata = getattr(cls, "metadata", None)
            metadata = cls.__fsa__._make_metadata(bind_key)

            if metadata is not parent_metadata:
                cls.metadata = metadata

        super().__init_subclass__(**kwargs)


class NameMetaMixin(type):
    """Metaclass mixin that sets a model's ``__tablename__`` by converting the
    ``CamelCase`` class name to ``snake_case``. A name is set for non-abstract models
    that do not otherwise define ``__tablename__``. If a model does not define a primary
    key, it will not generate a name or ``__table__``, for single-table inheritance.
    """

    metadata: sa.MetaData
    __tablename__: str
    __table__: sa.Table

    def __init__(
        cls, name: str, bases: tuple[type, ...], d: dict[str, t.Any], **kwargs: t.Any
    ) -> None:
        if should_set_tablename(cls):
            cls.__tablename__ = camel_to_snake_case(cls.__name__)

        super().__init__(name, bases, d, **kwargs)

        # __table_cls__ has run. If no table was created, use the parent table.
        if (
            "__tablename__" not in cls.__dict__
            and "__table__" in cls.__dict__
            and cls.__dict__["__table__"] is None
        ):
            del cls.__table__

    def __table_cls__(cls, *args: t.Any, **kwargs: t.Any) -> sa.Table | None:
        """This is called by SQLAlchemy during mapper setup. It determines the final
        table object that the model will use.

        If no primary key is found, that indicates single-table inheritance, so no table
        will be created and ``__tablename__`` will be unset.
        """
        schema = kwargs.get("schema")

        if schema is None:
            key = args[0]
        else:
            key = f"{schema}.{args[0]}"

        # Check if a table with this name already exists. Allows reflected tables to be
        # applied to models by name.
        if key in cls.metadata.tables:
            return sa.Table(*args, **kwargs)

        # If a primary key is found, create a table for joined-table inheritance.
        for arg in args:
            if (isinstance(arg, sa.Column) and arg.primary_key) or isinstance(
                arg, sa.PrimaryKeyConstraint
            ):
                return sa.Table(*args, **kwargs)

        # If no base classes define a table, return one that's missing a primary key
        # so SQLAlchemy shows the correct error.
        for base in cls.__mro__[1:-1]:
            if "__table__" in base.__dict__:
                break
        else:
            return sa.Table(*args, **kwargs)

        # Single-table inheritance, use the parent table name. __init__ will unset
        # __table__ based on this.
        if "__tablename__" in cls.__dict__:
            del cls.__tablename__

        return None


class NameMixin:
    """DeclarativeBase mixin that sets a model's ``__tablename__`` by converting the
    ``CamelCase`` class name to ``snake_case``. A name is set for non-abstract models
    that do not otherwise define ``__tablename__``. If a model does not define a primary
    key, it will not generate a name or ``__table__``, for single-table inheritance.

    .. versionchanged:: 3.1.0
    """

    metadata: sa.MetaData
    __tablename__: str
    __table__: sa.Table

    @classmethod
    def __init_subclass__(cls: t.Type[NameMixin], **kwargs: t.Dict[str, t.Any]) -> None:
        if should_set_tablename(cls):
            cls.__tablename__ = camel_to_snake_case(cls.__name__)

        super().__init_subclass__(**kwargs)

        # __table_cls__ has run. If no table was created, use the parent table.
        if (
            "__tablename__" not in cls.__dict__
            and "__table__" in cls.__dict__
            and cls.__dict__["__table__"] is None
        ):
            del cls.__table__

    @classmethod
    def __table_cls__(cls, *args: t.Any, **kwargs: t.Any) -> sa.Table | None:
        """This is called by SQLAlchemy during mapper setup. It determines the final
        table object that the model will use.

        If no primary key is found, that indicates single-table inheritance, so no table
        will be created and ``__tablename__`` will be unset.
        """
        schema = kwargs.get("schema")

        if schema is None:
            key = args[0]
        else:
            key = f"{schema}.{args[0]}"

        # Check if a table with this name already exists. Allows reflected tables to be
        # applied to models by name.
        if key in cls.metadata.tables:
            return sa.Table(*args, **kwargs)

        # If a primary key is found, create a table for joined-table inheritance.
        for arg in args:
            if (isinstance(arg, sa.Column) and arg.primary_key) or isinstance(
                arg, sa.PrimaryKeyConstraint
            ):
                return sa.Table(*args, **kwargs)

        # If no base classes define a table, return one that's missing a primary key
        # so SQLAlchemy shows the correct error.
        for base in cls.__mro__[1:-1]:
            if "__table__" in base.__dict__:
                break
        else:
            return sa.Table(*args, **kwargs)

        # Single-table inheritance, use the parent table name. __init__ will unset
        # __table__ based on this.
        if "__tablename__" in cls.__dict__:
            del cls.__tablename__

        return None


def should_set_tablename(cls: type) -> bool:
    """Determine whether ``__tablename__`` should be generated for a model.

    -   If no class in the MRO sets a name, one should be generated.
    -   If a declared attr is found, it should be used instead.
    -   If a name is found, it should be used if the class is a mixin, otherwise one
        should be generated.
    -   Abstract models should not have one generated.

    Later, ``__table_cls__`` will determine if the model looks like single or
    joined-table inheritance. If no primary key is found, the name will be unset.
    """
    if (
        cls.__dict__.get("__abstract__", False)
        or (
            not issubclass(cls, (sa_orm.DeclarativeBase, sa_orm.DeclarativeBaseNoMeta))
            and not any(isinstance(b, sa_orm.DeclarativeMeta) for b in cls.__mro__[1:])
        )
        or any(
            (b is sa_orm.DeclarativeBase or b is sa_orm.DeclarativeBaseNoMeta)
            for b in cls.__bases__
        )
    ):
        return False

    for base in cls.__mro__:
        if "__tablename__" not in base.__dict__:
            continue

        if isinstance(base.__dict__["__tablename__"], sa_orm.declared_attr):
            return False

        return not (
            base is cls
            or base.__dict__.get("__abstract__", False)
            or not (
                # SQLAlchemy 1.x
                isinstance(base, sa_orm.DeclarativeMeta)
                # 2.x: DeclarativeBas uses this as metaclass
                or isinstance(base, sa_orm.decl_api.DeclarativeAttributeIntercept)
                # 2.x: DeclarativeBaseNoMeta doesn't use a metaclass
                or issubclass(base, sa_orm.DeclarativeBaseNoMeta)
            )
        )

    return True


def camel_to_snake_case(name: str) -> str:
    """Convert a ``CamelCase`` name to ``snake_case``."""
    name = re.sub(r"((?<=[a-z0-9])[A-Z]|(?!^)[A-Z](?=[a-z]))", r"_\1", name)
    return name.lower().lstrip("_")


class DefaultMeta(BindMetaMixin, NameMetaMixin, sa_orm.DeclarativeMeta):
    """SQLAlchemy declarative metaclass that provides ``__bind_key__`` and
    ``__tablename__`` support.
    """


class DefaultMetaNoName(BindMetaMixin, sa_orm.DeclarativeMeta):
    """SQLAlchemy declarative metaclass that provides ``__bind_key__`` and
    ``__tablename__`` support.
    """

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\io\formats\csvs.py ===
"""
Module for formatting output data into CSV files.
"""

from __future__ import annotations

from collections.abc import (
    Hashable,
    Iterable,
    Iterator,
    Sequence,
)
import csv as csvlib
import os
from typing import (
    TYPE_CHECKING,
    Any,
    cast,
)

import numpy as np

from pandas._libs import writers as libwriters
from pandas._typing import SequenceNotStr
from pandas.util._decorators import cache_readonly

from pandas.core.dtypes.generic import (
    ABCDatetimeIndex,
    ABCIndex,
    ABCMultiIndex,
    ABCPeriodIndex,
)
from pandas.core.dtypes.missing import notna

from pandas.core.indexes.api import Index

from pandas.io.common import get_handle

if TYPE_CHECKING:
    from pandas._typing import (
        CompressionOptions,
        FilePath,
        FloatFormatType,
        IndexLabel,
        StorageOptions,
        WriteBuffer,
        npt,
    )

    from pandas.io.formats.format import DataFrameFormatter


_DEFAULT_CHUNKSIZE_CELLS = 100_000


class CSVFormatter:
    cols: npt.NDArray[np.object_]

    def __init__(
        self,
        formatter: DataFrameFormatter,
        path_or_buf: FilePath | WriteBuffer[str] | WriteBuffer[bytes] = "",
        sep: str = ",",
        cols: Sequence[Hashable] | None = None,
        index_label: IndexLabel | None = None,
        mode: str = "w",
        encoding: str | None = None,
        errors: str = "strict",
        compression: CompressionOptions = "infer",
        quoting: int | None = None,
        lineterminator: str | None = "\n",
        chunksize: int | None = None,
        quotechar: str | None = '"',
        date_format: str | None = None,
        doublequote: bool = True,
        escapechar: str | None = None,
        storage_options: StorageOptions | None = None,
    ) -> None:
        self.fmt = formatter

        self.obj = self.fmt.frame

        self.filepath_or_buffer = path_or_buf
        self.encoding = encoding
        self.compression: CompressionOptions = compression
        self.mode = mode
        self.storage_options = storage_options

        self.sep = sep
        self.index_label = self._initialize_index_label(index_label)
        self.errors = errors
        self.quoting = quoting or csvlib.QUOTE_MINIMAL
        self.quotechar = self._initialize_quotechar(quotechar)
        self.doublequote = doublequote
        self.escapechar = escapechar
        self.lineterminator = lineterminator or os.linesep
        self.date_format = date_format
        self.cols = self._initialize_columns(cols)
        self.chunksize = self._initialize_chunksize(chunksize)

    @property
    def na_rep(self) -> str:
        return self.fmt.na_rep

    @property
    def float_format(self) -> FloatFormatType | None:
        return self.fmt.float_format

    @property
    def decimal(self) -> str:
        return self.fmt.decimal

    @property
    def header(self) -> bool | SequenceNotStr[str]:
        return self.fmt.header

    @property
    def index(self) -> bool:
        return self.fmt.index

    def _initialize_index_label(self, index_label: IndexLabel | None) -> IndexLabel:
        if index_label is not False:
            if index_label is None:
                return self._get_index_label_from_obj()
            elif not isinstance(index_label, (list, tuple, np.ndarray, ABCIndex)):
                # given a string for a DF with Index
                return [index_label]
        return index_label

    def _get_index_label_from_obj(self) -> Sequence[Hashable]:
        if isinstance(self.obj.index, ABCMultiIndex):
            return self._get_index_label_multiindex()
        else:
            return self._get_index_label_flat()

    def _get_index_label_multiindex(self) -> Sequence[Hashable]:
        return [name or "" for name in self.obj.index.names]

    def _get_index_label_flat(self) -> Sequence[Hashable]:
        index_label = self.obj.index.name
        return [""] if index_label is None else [index_label]

    def _initialize_quotechar(self, quotechar: str | None) -> str | None:
        if self.quoting != csvlib.QUOTE_NONE:
            # prevents crash in _csv
            return quotechar
        return None

    @property
    def has_mi_columns(self) -> bool:
        return bool(isinstance(self.obj.columns, ABCMultiIndex))

    def _initialize_columns(
        self, cols: Iterable[Hashable] | None
    ) -> npt.NDArray[np.object_]:
        # validate mi options
        if self.has_mi_columns:
            if cols is not None:
                msg = "cannot specify cols with a MultiIndex on the columns"
                raise TypeError(msg)

        if cols is not None:
            if isinstance(cols, ABCIndex):
                cols = cols._get_values_for_csv(**self._number_format)
            else:
                cols = list(cols)
            self.obj = self.obj.loc[:, cols]

        # update columns to include possible multiplicity of dupes
        # and make sure cols is just a list of labels
        new_cols = self.obj.columns
        return new_cols._get_values_for_csv(**self._number_format)

    def _initialize_chunksize(self, chunksize: int | None) -> int:
        if chunksize is None:
            return (_DEFAULT_CHUNKSIZE_CELLS // (len(self.cols) or 1)) or 1
        return int(chunksize)

    @property
    def _number_format(self) -> dict[str, Any]:
        """Dictionary used for storing number formatting settings."""
        return {
            "na_rep": self.na_rep,
            "float_format": self.float_format,
            "date_format": self.date_format,
            "quoting": self.quoting,
            "decimal": self.decimal,
        }

    @cache_readonly
    def data_index(self) -> Index:
        data_index = self.obj.index
        if (
            isinstance(data_index, (ABCDatetimeIndex, ABCPeriodIndex))
            and self.date_format is not None
        ):
            data_index = Index(
                [x.strftime(self.date_format) if notna(x) else "" for x in data_index]
            )
        elif isinstance(data_index, ABCMultiIndex):
            data_index = data_index.remove_unused_levels()
        return data_index

    @property
    def nlevels(self) -> int:
        if self.index:
            return getattr(self.data_index, "nlevels", 1)
        else:
            return 0

    @property
    def _has_aliases(self) -> bool:
        return isinstance(self.header, (tuple, list, np.ndarray, ABCIndex))

    @property
    def _need_to_save_header(self) -> bool:
        return bool(self._has_aliases or self.header)

    @property
    def write_cols(self) -> SequenceNotStr[Hashable]:
        if self._has_aliases:
            assert not isinstance(self.header, bool)
            if len(self.header) != len(self.cols):
                raise ValueError(
                    f"Writing {len(self.cols)} cols but got {len(self.header)} aliases"
                )
            return self.header
        else:
            # self.cols is an ndarray derived from Index._get_values_for_csv,
            #  so its entries are strings, i.e. hashable
            return cast(SequenceNotStr[Hashable], self.cols)

    @property
    def encoded_labels(self) -> list[Hashable]:
        encoded_labels: list[Hashable] = []

        if self.index and self.index_label:
            assert isinstance(self.index_label, Sequence)
            encoded_labels = list(self.index_label)

        if not self.has_mi_columns or self._has_aliases:
            encoded_labels += list(self.write_cols)

        return encoded_labels

    def save(self) -> None:
        """
        Create the writer & save.
        """
        # apply compression and byte/text conversion
        with get_handle(
            self.filepath_or_buffer,
            self.mode,
            encoding=self.encoding,
            errors=self.errors,
            compression=self.compression,
            storage_options=self.storage_options,
        ) as handles:
            # Note: self.encoding is irrelevant here
            self.writer = csvlib.writer(
                handles.handle,
                lineterminator=self.lineterminator,
                delimiter=self.sep,
                quoting=self.quoting,
                doublequote=self.doublequote,
                escapechar=self.escapechar,
                quotechar=self.quotechar,
            )

            self._save()

    def _save(self) -> None:
        if self._need_to_save_header:
            self._save_header()
        self._save_body()

    def _save_header(self) -> None:
        if not self.has_mi_columns or self._has_aliases:
            self.writer.writerow(self.encoded_labels)
        else:
            for row in self._generate_multiindex_header_rows():
                self.writer.writerow(row)

    def _generate_multiindex_header_rows(self) -> Iterator[list[Hashable]]:
        columns = self.obj.columns
        for i in range(columns.nlevels):
            # we need at least 1 index column to write our col names
            col_line = []
            if self.index:
                # name is the first column
                col_line.append(columns.names[i])

                if isinstance(self.index_label, list) and len(self.index_label) > 1:
                    col_line.extend([""] * (len(self.index_label) - 1))

            col_line.extend(columns._get_level_values(i))
            yield col_line

        # Write out the index line if it's not empty.
        # Otherwise, we will print out an extraneous
        # blank line between the mi and the data rows.
        if self.encoded_labels and set(self.encoded_labels) != {""}:
            yield self.encoded_labels + [""] * len(columns)

    def _save_body(self) -> None:
        nrows = len(self.data_index)
        chunks = (nrows // self.chunksize) + 1
        for i in range(chunks):
            start_i = i * self.chunksize
            end_i = min(start_i + self.chunksize, nrows)
            if start_i >= end_i:
                break
            self._save_chunk(start_i, end_i)

    def _save_chunk(self, start_i: int, end_i: int) -> None:
        # create the data for a chunk
        slicer = slice(start_i, end_i)
        df = self.obj.iloc[slicer]

        res = df._get_values_for_csv(**self._number_format)
        data = list(res._iter_column_arrays())

        ix = self.data_index[slicer]._get_values_for_csv(**self._number_format)
        libwriters.write_csv_rows(
            data,
            ix,
            self.nlevels,
            self.cols,
            self.writer,
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\firefox\firefox_profile.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

import base64
import copy
import json
import os
import re
import shutil
import sys
import tempfile
import warnings
import zipfile
from io import BytesIO
from xml.dom import minidom

from typing_extensions import deprecated

from selenium.common.exceptions import WebDriverException

WEBDRIVER_PREFERENCES = "webdriver_prefs.json"


@deprecated("Addons must be added after starting the session")
class AddonFormatError(Exception):
    """Exception for not well-formed add-on manifest files."""


class FirefoxProfile:
    DEFAULT_PREFERENCES = None

    def __init__(self, profile_directory=None):
        """Initialises a new instance of a Firefox Profile.

        :args:
         - profile_directory: Directory of profile that you want to use. If a
           directory is passed in it will be cloned and the cloned directory
           will be used by the driver when instantiated.
           This defaults to None and will create a new
           directory when object is created.
        """
        self._desired_preferences = {}
        if profile_directory:
            newprof = os.path.join(tempfile.mkdtemp(), "webdriver-py-profilecopy")
            shutil.copytree(
                profile_directory, newprof, ignore=shutil.ignore_patterns("parent.lock", "lock", ".parentlock")
            )
            self._profile_dir = newprof
            os.chmod(self._profile_dir, 0o755)
        else:
            self._profile_dir = tempfile.mkdtemp()
            if not FirefoxProfile.DEFAULT_PREFERENCES:
                with open(
                    os.path.join(os.path.dirname(__file__), WEBDRIVER_PREFERENCES), encoding="utf-8"
                ) as default_prefs:
                    FirefoxProfile.DEFAULT_PREFERENCES = json.load(default_prefs)

            self._desired_preferences = copy.deepcopy(FirefoxProfile.DEFAULT_PREFERENCES["mutable"])
            for key, value in FirefoxProfile.DEFAULT_PREFERENCES["frozen"].items():
                self._desired_preferences[key] = value

    # Public Methods
    def set_preference(self, key, value):
        """Sets the preference that we want in the profile."""
        self._desired_preferences[key] = value

    @deprecated("Addons must be added after starting the session")
    def add_extension(self, extension=None):
        self._install_extension(extension)

    def update_preferences(self):
        """Writes the desired user prefs to disk."""
        user_prefs = os.path.join(self._profile_dir, "user.js")
        if os.path.isfile(user_prefs):
            os.chmod(user_prefs, 0o644)
            self._read_existing_userjs(user_prefs)
        with open(user_prefs, "w", encoding="utf-8") as f:
            for key, value in self._desired_preferences.items():
                f.write(f'user_pref("{key}", {json.dumps(value)});\n')

    # Properties

    @property
    def path(self):
        """Gets the profile directory that is currently being used."""
        return self._profile_dir

    @property
    @deprecated("The port is stored in the Service class")
    def port(self):
        """Gets the port that WebDriver is working on."""
        return self._port

    @port.setter
    @deprecated("The port is stored in the Service class")
    def port(self, port) -> None:
        """Sets the port that WebDriver will be running on."""
        if not isinstance(port, int):
            raise WebDriverException("Port needs to be an integer")
        try:
            port = int(port)
            if port < 1 or port > 65535:
                raise WebDriverException("Port number must be in the range 1..65535")
        except (ValueError, TypeError):
            raise WebDriverException("Port needs to be an integer")
        self._port = port
        self.set_preference("webdriver_firefox_port", self._port)

    @property
    @deprecated("Allowing untrusted certs is toggled in the Options class")
    def accept_untrusted_certs(self):
        return self._desired_preferences["webdriver_accept_untrusted_certs"]

    @accept_untrusted_certs.setter
    @deprecated("Allowing untrusted certs is toggled in the Options class")
    def accept_untrusted_certs(self, value) -> None:
        if not isinstance(value, bool):
            raise WebDriverException("Please pass in a Boolean to this call")
        self.set_preference("webdriver_accept_untrusted_certs", value)

    @property
    @deprecated("Allowing untrusted certs is toggled in the Options class")
    def assume_untrusted_cert_issuer(self):
        return self._desired_preferences["webdriver_assume_untrusted_issuer"]

    @assume_untrusted_cert_issuer.setter
    @deprecated("Allowing untrusted certs is toggled in the Options class")
    def assume_untrusted_cert_issuer(self, value) -> None:
        if not isinstance(value, bool):
            raise WebDriverException("Please pass in a Boolean to this call")

        self.set_preference("webdriver_assume_untrusted_issuer", value)

    @property
    def encoded(self) -> str:
        """Updates preferences and creates a zipped, base64 encoded string of
        profile directory."""
        if self._desired_preferences:
            self.update_preferences()
        fp = BytesIO()
        with zipfile.ZipFile(fp, "w", zipfile.ZIP_DEFLATED, strict_timestamps=False) as zipped:
            path_root = len(self.path) + 1  # account for trailing slash
            for base, _, files in os.walk(self.path):
                for fyle in files:
                    filename = os.path.join(base, fyle)
                    zipped.write(filename, filename[path_root:])
        return base64.b64encode(fp.getvalue()).decode("UTF-8")

    def _read_existing_userjs(self, userjs):
        """Reads existing preferences and adds them to desired preference
        dictionary."""
        pref_pattern = re.compile(r'user_pref\("(.*)",\s(.*)\)')
        with open(userjs, encoding="utf-8") as f:
            for usr in f:
                matches = pref_pattern.search(usr)
                try:
                    self._desired_preferences[matches.group(1)] = json.loads(matches.group(2))
                except Exception:
                    warnings.warn(
                        f"(skipping) failed to json.loads existing preference: {matches.group(1) + matches.group(2)}"
                    )

    @deprecated("Addons must be added after starting the session")
    def _install_extension(self, addon, unpack=True):
        """Installs addon from a filepath, url or directory of addons in the
        profile.

        - path: url, absolute path to .xpi, or directory of addons
        - unpack: whether to unpack unless specified otherwise in the install.rdf
        """
        tmpdir = None
        xpifile = None
        if addon.endswith(".xpi"):
            tmpdir = tempfile.mkdtemp(suffix="." + os.path.split(addon)[-1])
            compressed_file = zipfile.ZipFile(addon, "r")
            for name in compressed_file.namelist():
                if name.endswith("/"):
                    if not os.path.isdir(os.path.join(tmpdir, name)):
                        os.makedirs(os.path.join(tmpdir, name))
                else:
                    if not os.path.isdir(os.path.dirname(os.path.join(tmpdir, name))):
                        os.makedirs(os.path.dirname(os.path.join(tmpdir, name)))
                    data = compressed_file.read(name)
                    with open(os.path.join(tmpdir, name), "wb") as f:
                        f.write(data)
            xpifile = addon
            addon = tmpdir

        # determine the addon id
        addon_details = self._addon_details(addon)
        addon_id = addon_details.get("id")
        assert addon_id, f"The addon id could not be found: {addon}"

        # copy the addon to the profile
        extensions_dir = os.path.join(self._profile_dir, "extensions")
        addon_path = os.path.join(extensions_dir, addon_id)
        if not unpack and not addon_details["unpack"] and xpifile:
            if not os.path.exists(extensions_dir):
                os.makedirs(extensions_dir)
                os.chmod(extensions_dir, 0o755)
            shutil.copy(xpifile, addon_path + ".xpi")
        else:
            if not os.path.exists(addon_path):
                shutil.copytree(addon, addon_path, symlinks=True)

        # remove the temporary directory, if any
        if tmpdir:
            shutil.rmtree(tmpdir)

    @deprecated("Addons must be added after starting the session")
    def _addon_details(self, addon_path):
        """Returns a dictionary of details about the addon.

        :param addon_path: path to the add-on directory or XPI

        Returns::

            {
                "id": "rainbow@colors.org",  # id of the addon
                "version": "1.4",  # version of the addon
                "name": "Rainbow",  # name of the addon
                "unpack": False,
            }  # whether to unpack the addon
        """

        details = {"id": None, "unpack": False, "name": None, "version": None}

        def get_namespace_id(doc, url):
            attributes = doc.documentElement.attributes
            namespace = ""
            for i in range(attributes.length):
                if attributes.item(i).value == url:
                    if ":" in attributes.item(i).name:
                        # If the namespace is not the default one remove 'xlmns:'
                        namespace = attributes.item(i).name.split(":")[1] + ":"
                        break
            return namespace

        def get_text(element):
            """Retrieve the text value of a given node."""
            rc = []
            for node in element.childNodes:
                if node.nodeType == node.TEXT_NODE:
                    rc.append(node.data)
            return "".join(rc).strip()

        def parse_manifest_json(content):
            """Extracts the details from the contents of a WebExtensions
            `manifest.json` file."""
            manifest = json.loads(content)
            try:
                id = manifest["applications"]["gecko"]["id"]
            except KeyError:
                id = manifest["name"].replace(" ", "") + "@" + manifest["version"]
            return {
                "id": id,
                "version": manifest["version"],
                "name": manifest["version"],
                "unpack": False,
            }

        if not os.path.exists(addon_path):
            raise OSError(f"Add-on path does not exist: {addon_path}")

        try:
            if zipfile.is_zipfile(addon_path):
                with zipfile.ZipFile(addon_path, "r") as compressed_file:
                    if "manifest.json" in compressed_file.namelist():
                        return parse_manifest_json(compressed_file.read("manifest.json"))

                    manifest = compressed_file.read("install.rdf")
            elif os.path.isdir(addon_path):
                manifest_json_filename = os.path.join(addon_path, "manifest.json")
                if os.path.exists(manifest_json_filename):
                    with open(manifest_json_filename, encoding="utf-8") as f:
                        return parse_manifest_json(f.read())

                with open(os.path.join(addon_path, "install.rdf"), encoding="utf-8") as f:
                    manifest = f.read()
            else:
                raise OSError(f"Add-on path is neither an XPI nor a directory: {addon_path}")
        except (OSError, KeyError) as e:
            raise AddonFormatError(str(e), sys.exc_info()[2])

        try:
            doc = minidom.parseString(manifest)

            # Get the namespaces abbreviations
            em = get_namespace_id(doc, "http://www.mozilla.org/2004/em-rdf#")
            rdf = get_namespace_id(doc, "http://www.w3.org/1999/02/22-rdf-syntax-ns#")

            description = doc.getElementsByTagName(rdf + "Description").item(0)
            if not description:
                description = doc.getElementsByTagName("Description").item(0)
            for node in description.childNodes:
                # Remove the namespace prefix from the tag for comparison
                entry = node.nodeName.replace(em, "")
                if entry in details:
                    details.update({entry: get_text(node)})
            if not details.get("id"):
                for i in range(description.attributes.length):
                    attribute = description.attributes.item(i)
                    if attribute.name == em + "id":
                        details.update({"id": attribute.value})
        except Exception as e:
            raise AddonFormatError(str(e), sys.exc_info()[2])

        # turn unpack into a true/false value
        if isinstance(details["unpack"], str):
            details["unpack"] = details["unpack"].lower() == "true"

        # If no ID is set, the add-on is invalid
        if not details.get("id"):
            raise AddonFormatError("Add-on id could not be found.")

        return details

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\wtforms\form.py ===
import itertools
from collections import OrderedDict

from wtforms.meta import DefaultMeta
from wtforms.utils import unset_value

__all__ = ("BaseForm", "Form")

_default_meta = DefaultMeta()


class BaseForm:
    """
    Base Form Class.  Provides core behaviour like field construction,
    validation, and data and error proxying.
    """

    def __init__(self, fields, prefix="", meta=_default_meta):
        """
        :param fields:
            A dict or sequence of 2-tuples of partially-constructed fields.
        :param prefix:
            If provided, all fields will have their name prefixed with the
            value.
        :param meta:
            A meta instance which is used for configuration and customization
            of WTForms behaviors.
        """
        if prefix and prefix[-1] not in "-_;:/.":
            prefix += "-"

        self.meta = meta
        self._form_error_key = ""
        self._prefix = prefix
        self._fields = OrderedDict()

        if hasattr(fields, "items"):
            fields = fields.items()

        translations = self.meta.get_translations(self)
        extra_fields = []
        if meta.csrf:
            self._csrf = meta.build_csrf(self)
            extra_fields.extend(self._csrf.setup_form(self))

        for name, unbound_field in itertools.chain(fields, extra_fields):
            field_name = unbound_field.name or name
            options = dict(name=field_name, prefix=prefix, translations=translations)
            field = meta.bind_field(self, unbound_field, options)
            self._fields[name] = field

        self.form_errors = []

    def __iter__(self):
        """Iterate form fields in creation order."""
        return iter(self._fields.values())

    def __contains__(self, name):
        """Returns `True` if the named field is a member of this form."""
        return name in self._fields

    def __getitem__(self, name):
        """Dict-style access to this form's fields."""
        return self._fields[name]

    def __setitem__(self, name, value):
        """Bind a field to this form."""
        self._fields[name] = value.bind(form=self, name=name, prefix=self._prefix)

    def __delitem__(self, name):
        """Remove a field from this form."""
        del self._fields[name]

    def populate_obj(self, obj):
        """
        Populates the attributes of the passed `obj` with data from the form's
        fields.

        :note: This is a destructive operation; Any attribute with the same name
               as a field will be overridden. Use with caution.
        """
        for name, field in self._fields.items():
            field.populate_obj(obj, name)

    def process(self, formdata=None, obj=None, data=None, extra_filters=None, **kwargs):
        """Process default and input data with each field.

        :param formdata: Input data coming from the client, usually
            ``request.form`` or equivalent. Should provide a "multi
            dict" interface to get a list of values for a given key,
            such as what Werkzeug, Django, and WebOb provide.
        :param obj: Take existing data from attributes on this object
            matching form field attributes. Only used if ``formdata`` is
            not passed.
        :param data: Take existing data from keys in this dict matching
            form field attributes. ``obj`` takes precedence if it also
            has a matching attribute. Only used if ``formdata`` is not
            passed.
        :param extra_filters: A dict mapping field attribute names to
            lists of extra filter functions to run. Extra filters run
            after filters passed when creating the field. If the form
            has ``filter_<fieldname>``, it is the last extra filter.
        :param kwargs: Merged with ``data`` to allow passing existing
            data as parameters. Overwrites any duplicate keys in
            ``data``. Only used if ``formdata`` is not passed.
        """
        formdata = self.meta.wrap_formdata(self, formdata)

        if data is not None:
            kwargs = dict(data, **kwargs)

        filters = extra_filters.copy() if extra_filters is not None else {}

        for name, field in self._fields.items():
            field_extra_filters = filters.get(name, [])

            inline_filter = getattr(self, f"filter_{name}", None)
            if inline_filter is not None:
                field_extra_filters.append(inline_filter)

            if obj is not None and hasattr(obj, name):
                data = getattr(obj, name)
            elif name in kwargs:
                data = kwargs[name]
            else:
                data = unset_value

            field.process(formdata, data, extra_filters=field_extra_filters)

    def validate(self, extra_validators=None):
        """
        Validates the form by calling `validate` on each field.

        :param extra_validators:
            If provided, is a dict mapping field names to a sequence of
            callables which will be passed as extra validators to the field's
            `validate` method.

        Returns `True` if no errors occur.
        """
        success = True
        for name, field in self._fields.items():
            if extra_validators is not None and name in extra_validators:
                extra = extra_validators[name]
            else:
                extra = tuple()
            if not field.validate(self, extra):
                success = False
        return success

    @property
    def data(self):
        return {name: f.data for name, f in self._fields.items()}

    @property
    def errors(self):
        errors = {name: f.errors for name, f in self._fields.items() if f.errors}
        if self.form_errors:
            errors[self._form_error_key] = self.form_errors
        return errors


class FormMeta(type):
    """
    The metaclass for `Form` and any subclasses of `Form`.

    `FormMeta`'s responsibility is to create the `_unbound_fields` list, which
    is a list of `UnboundField` instances sorted by their order of
    instantiation.  The list is created at the first instantiation of the form.
    If any fields are added/removed from the form, the list is cleared to be
    re-generated on the next instantiation.

    Any properties which begin with an underscore or are not `UnboundField`
    instances are ignored by the metaclass.
    """

    def __init__(cls, name, bases, attrs):
        type.__init__(cls, name, bases, attrs)
        cls._unbound_fields = None
        cls._wtforms_meta = None

    def __call__(cls, *args, **kwargs):
        """
        Construct a new `Form` instance.

        Creates the `_unbound_fields` list and the internal `_wtforms_meta`
        subclass of the class Meta in order to allow a proper inheritance
        hierarchy.
        """
        if cls._unbound_fields is None:
            fields = []
            for name in dir(cls):
                if not name.startswith("_"):
                    unbound_field = getattr(cls, name)
                    if hasattr(unbound_field, "_formfield"):
                        fields.append((name, unbound_field))
            # We keep the name as the second element of the sort
            # to ensure a stable sort.
            fields.sort(key=lambda x: (x[1].creation_counter, x[0]))
            cls._unbound_fields = fields

        # Create a subclass of the 'class Meta' using all the ancestors.
        if cls._wtforms_meta is None:
            bases = []
            for mro_class in cls.__mro__:
                if "Meta" in mro_class.__dict__:
                    bases.append(mro_class.Meta)
            cls._wtforms_meta = type("Meta", tuple(bases), {})
        return type.__call__(cls, *args, **kwargs)

    def __setattr__(cls, name, value):
        """
        Add an attribute to the class, clearing `_unbound_fields` if needed.
        """
        if name == "Meta":
            cls._wtforms_meta = None
        elif not name.startswith("_") and hasattr(value, "_formfield"):
            cls._unbound_fields = None
        type.__setattr__(cls, name, value)

    def __delattr__(cls, name):
        """
        Remove an attribute from the class, clearing `_unbound_fields` if
        needed.
        """
        if not name.startswith("_"):
            cls._unbound_fields = None
        type.__delattr__(cls, name)


class Form(BaseForm, metaclass=FormMeta):
    """
    Declarative Form base class. Extends BaseForm's core behaviour allowing
    fields to be defined on Form subclasses as class attributes.

    In addition, form and instance input data are taken at construction time
    and passed to `process()`.
    """

    Meta = DefaultMeta

    def __init__(
        self,
        formdata=None,
        obj=None,
        prefix="",
        data=None,
        meta=None,
        **kwargs,
    ):
        """
        :param formdata: Input data coming from the client, usually
            ``request.form`` or equivalent. Should provide a "multi
            dict" interface to get a list of values for a given key,
            such as what Werkzeug, Django, and WebOb provide.
        :param obj: Take existing data from attributes on this object
            matching form field attributes. Only used if ``formdata`` is
            not passed.
        :param prefix: If provided, all fields will have their name
            prefixed with the value. This is for distinguishing multiple
            forms on a single page. This only affects the HTML name for
            matching input data, not the Python name for matching
            existing data.
        :param data: Take existing data from keys in this dict matching
            form field attributes. ``obj`` takes precedence if it also
            has a matching attribute. Only used if ``formdata`` is not
            passed.
        :param meta: A dict of attributes to override on this form's
            :attr:`meta` instance.
        :param extra_filters: A dict mapping field attribute names to
            lists of extra filter functions to run. Extra filters run
            after filters passed when creating the field. If the form
            has ``filter_<fieldname>``, it is the last extra filter.
        :param kwargs: Merged with ``data`` to allow passing existing
            data as parameters. Overwrites any duplicate keys in
            ``data``. Only used if ``formdata`` is not passed.
        """
        meta_obj = self._wtforms_meta()
        if meta is not None and isinstance(meta, dict):
            meta_obj.update_values(meta)
        super().__init__(self._unbound_fields, meta=meta_obj, prefix=prefix)

        for name, field in self._fields.items():
            # Set all the fields to attributes so that they obscure the class
            # attributes with the same names.
            setattr(self, name, field)
        self.process(formdata, obj, data=data, **kwargs)

    def __setitem__(self, name, value):
        raise TypeError("Fields may not be added to Form instances, only classes.")

    def __delitem__(self, name):
        del self._fields[name]
        setattr(self, name, None)

    def __delattr__(self, name):
        if name in self._fields:
            self.__delitem__(name)
        else:
            # This is done for idempotency, if we have a name which is a field,
            # we want to mask it by setting the value to None.
            unbound_field = getattr(self.__class__, name, None)
            if unbound_field is not None and hasattr(unbound_field, "_formfield"):
                setattr(self, name, None)
            else:
                super().__delattr__(name)

    def validate(self, extra_validators=None):
        """Validate the form by calling ``validate`` on each field.
        Returns ``True`` if validation passes.

        If the form defines a ``validate_<fieldname>`` method, it is
        appended as an extra validator for the field's ``validate``.

        :param extra_validators: A dict mapping field names to lists of
            extra validator methods to run. Extra validators run after
            validators passed when creating the field. If the form has
            ``validate_<fieldname>``, it is the last extra validator.
        """
        if extra_validators is not None:
            extra = extra_validators.copy()
        else:
            extra = {}

        for name in self._fields:
            inline = getattr(self.__class__, f"validate_{name}", None)
            if inline is not None:
                extra.setdefault(name, []).append(inline)

        return super().validate(extra)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\flask_wtf\csrf.py ===
import hashlib
import hmac
import logging
import os
from urllib.parse import urlparse

from flask import Blueprint
from flask import current_app
from flask import g
from flask import request
from flask import session
from itsdangerous import BadData
from itsdangerous import SignatureExpired
from itsdangerous import URLSafeTimedSerializer
from werkzeug.exceptions import BadRequest
from wtforms import ValidationError
from wtforms.csrf.core import CSRF

__all__ = ("generate_csrf", "validate_csrf", "CSRFProtect")
logger = logging.getLogger(__name__)


def generate_csrf(secret_key=None, token_key=None):
    """Generate a CSRF token. The token is cached for a request, so multiple
    calls to this function will generate the same token.

    During testing, it might be useful to access the signed token in
    ``g.csrf_token`` and the raw token in ``session['csrf_token']``.

    :param secret_key: Used to securely sign the token. Default is
        ``WTF_CSRF_SECRET_KEY`` or ``SECRET_KEY``.
    :param token_key: Key where token is stored in session for comparison.
        Default is ``WTF_CSRF_FIELD_NAME`` or ``'csrf_token'``.
    """

    secret_key = _get_config(
        secret_key,
        "WTF_CSRF_SECRET_KEY",
        current_app.secret_key,
        message="A secret key is required to use CSRF.",
    )
    field_name = _get_config(
        token_key,
        "WTF_CSRF_FIELD_NAME",
        "csrf_token",
        message="A field name is required to use CSRF.",
    )

    if field_name not in g:
        s = URLSafeTimedSerializer(secret_key, salt="wtf-csrf-token")

        if field_name not in session:
            session[field_name] = hashlib.sha1(os.urandom(64)).hexdigest()

        try:
            token = s.dumps(session[field_name])
        except TypeError:
            session[field_name] = hashlib.sha1(os.urandom(64)).hexdigest()
            token = s.dumps(session[field_name])

        setattr(g, field_name, token)

    return g.get(field_name)


def validate_csrf(data, secret_key=None, time_limit=None, token_key=None):
    """Check if the given data is a valid CSRF token. This compares the given
    signed token to the one stored in the session.

    :param data: The signed CSRF token to be checked.
    :param secret_key: Used to securely sign the token. Default is
        ``WTF_CSRF_SECRET_KEY`` or ``SECRET_KEY``.
    :param time_limit: Number of seconds that the token is valid. Default is
        ``WTF_CSRF_TIME_LIMIT`` or 3600 seconds (60 minutes).
    :param token_key: Key where token is stored in session for comparison.
        Default is ``WTF_CSRF_FIELD_NAME`` or ``'csrf_token'``.

    :raises ValidationError: Contains the reason that validation failed.

    .. versionchanged:: 0.14
        Raises ``ValidationError`` with a specific error message rather than
        returning ``True`` or ``False``.
    """

    secret_key = _get_config(
        secret_key,
        "WTF_CSRF_SECRET_KEY",
        current_app.secret_key,
        message="A secret key is required to use CSRF.",
    )
    field_name = _get_config(
        token_key,
        "WTF_CSRF_FIELD_NAME",
        "csrf_token",
        message="A field name is required to use CSRF.",
    )
    time_limit = _get_config(time_limit, "WTF_CSRF_TIME_LIMIT", 3600, required=False)

    if not data:
        raise ValidationError("The CSRF token is missing.")

    if field_name not in session:
        raise ValidationError("The CSRF session token is missing.")

    s = URLSafeTimedSerializer(secret_key, salt="wtf-csrf-token")

    try:
        token = s.loads(data, max_age=time_limit)
    except SignatureExpired as e:
        raise ValidationError("The CSRF token has expired.") from e
    except BadData as e:
        raise ValidationError("The CSRF token is invalid.") from e

    if not hmac.compare_digest(session[field_name], token):
        raise ValidationError("The CSRF tokens do not match.")


def _get_config(
    value, config_name, default=None, required=True, message="CSRF is not configured."
):
    """Find config value based on provided value, Flask config, and default
    value.

    :param value: already provided config value
    :param config_name: Flask ``config`` key
    :param default: default value if not provided or configured
    :param required: whether the value must not be ``None``
    :param message: error message if required config is not found
    :raises KeyError: if required config is not found
    """

    if value is None:
        value = current_app.config.get(config_name, default)

    if required and value is None:
        raise RuntimeError(message)

    return value


class _FlaskFormCSRF(CSRF):
    def setup_form(self, form):
        self.meta = form.meta
        return super().setup_form(form)

    def generate_csrf_token(self, csrf_token_field):
        return generate_csrf(
            secret_key=self.meta.csrf_secret, token_key=self.meta.csrf_field_name
        )

    def validate_csrf_token(self, form, field):
        if g.get("csrf_valid", False):
            # already validated by CSRFProtect
            return

        try:
            validate_csrf(
                field.data,
                self.meta.csrf_secret,
                self.meta.csrf_time_limit,
                self.meta.csrf_field_name,
            )
        except ValidationError as e:
            logger.info(e.args[0])
            raise


class CSRFProtect:
    """Enable CSRF protection globally for a Flask app.

    ::

        app = Flask(__name__)
        csrf = CSRFProtect(app)

    Checks the ``csrf_token`` field sent with forms, or the ``X-CSRFToken``
    header sent with JavaScript requests. Render the token in templates using
    ``{{ csrf_token() }}``.

    See the :ref:`csrf` documentation.
    """

    def __init__(self, app=None):
        self._exempt_views = set()
        self._exempt_blueprints = set()

        if app:
            self.init_app(app)

    def init_app(self, app):
        app.extensions["csrf"] = self

        app.config.setdefault("WTF_CSRF_ENABLED", True)
        app.config.setdefault("WTF_CSRF_CHECK_DEFAULT", True)
        app.config["WTF_CSRF_METHODS"] = set(
            app.config.get("WTF_CSRF_METHODS", ["POST", "PUT", "PATCH", "DELETE"])
        )
        app.config.setdefault("WTF_CSRF_FIELD_NAME", "csrf_token")
        app.config.setdefault("WTF_CSRF_HEADERS", ["X-CSRFToken", "X-CSRF-Token"])
        app.config.setdefault("WTF_CSRF_TIME_LIMIT", 3600)
        app.config.setdefault("WTF_CSRF_SSL_STRICT", True)

        app.jinja_env.globals["csrf_token"] = generate_csrf
        app.context_processor(lambda: {"csrf_token": generate_csrf})

        @app.before_request
        def csrf_protect():
            if not app.config["WTF_CSRF_ENABLED"]:
                return

            if not app.config["WTF_CSRF_CHECK_DEFAULT"]:
                return

            if request.method not in app.config["WTF_CSRF_METHODS"]:
                return

            if not request.endpoint:
                return

            if app.blueprints.get(request.blueprint) in self._exempt_blueprints:
                return

            view = app.view_functions.get(request.endpoint)
            dest = f"{view.__module__}.{view.__name__}"

            if dest in self._exempt_views:
                return

            self.protect()

    def _get_csrf_token(self):
        # find the token in the form data
        field_name = current_app.config["WTF_CSRF_FIELD_NAME"]
        base_token = request.form.get(field_name)

        if base_token:
            return base_token

        # if the form has a prefix, the name will be {prefix}-csrf_token
        for key in request.form:
            if key.endswith(field_name):
                csrf_token = request.form[key]

                if csrf_token:
                    return csrf_token

        # find the token in the headers
        for header_name in current_app.config["WTF_CSRF_HEADERS"]:
            csrf_token = request.headers.get(header_name)

            if csrf_token:
                return csrf_token

        return None

    def protect(self):
        if request.method not in current_app.config["WTF_CSRF_METHODS"]:
            return

        try:
            validate_csrf(self._get_csrf_token())
        except ValidationError as e:
            logger.info(e.args[0])
            self._error_response(e.args[0])

        if request.is_secure and current_app.config["WTF_CSRF_SSL_STRICT"]:
            if not request.referrer:
                self._error_response("The referrer header is missing.")

            good_referrer = f"https://{request.host}/"

            if not same_origin(request.referrer, good_referrer):
                self._error_response("The referrer does not match the host.")

        g.csrf_valid = True  # mark this request as CSRF valid

    def exempt(self, view):
        """Mark a view or blueprint to be excluded from CSRF protection.

        ::

            @app.route('/some-view', methods=['POST'])
            @csrf.exempt
            def some_view():
                ...

        ::

            bp = Blueprint(...)
            csrf.exempt(bp)

        """

        if isinstance(view, Blueprint):
            self._exempt_blueprints.add(view)
            return view

        if isinstance(view, str):
            view_location = view
        else:
            view_location = ".".join((view.__module__, view.__name__))

        self._exempt_views.add(view_location)
        return view

    def _error_response(self, reason):
        raise CSRFError(reason)


class CSRFError(BadRequest):
    """Raise if the client sends invalid CSRF data with the request.

    Generates a 400 Bad Request response with the failure reason by default.
    Customize the response by registering a handler with
    :meth:`flask.Flask.errorhandler`.
    """

    description = "CSRF validation failed."


def same_origin(current_uri, compare_uri):
    current = urlparse(current_uri)
    compare = urlparse(compare_uri)

    return (
        current.scheme == compare.scheme
        and current.hostname == compare.hostname
        and current.port == compare.port
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\f2py\func2subr.py ===
"""

Rules for building C/API module with f2py2e.

Copyright 1999 -- 2011 Pearu Peterson all rights reserved.
Copyright 2011 -- present NumPy Developers.
Permission to use, modify, and distribute this software is given under the
terms of the NumPy License.

NO WARRANTY IS EXPRESSED OR IMPLIED.  USE AT YOUR OWN RISK.
"""
import copy

from ._isocbind import isoc_kindmap
from .auxfuncs import (
    getfortranname,
    isexternal,
    isfunction,
    isfunction_wrap,
    isintent_in,
    isintent_out,
    islogicalfunction,
    ismoduleroutine,
    isscalar,
    issubroutine,
    issubroutine_wrap,
    outmess,
    show,
)


def var2fixfortran(vars, a, fa=None, f90mode=None):
    if fa is None:
        fa = a
    if a not in vars:
        show(vars)
        outmess(f'var2fixfortran: No definition for argument "{a}".\n')
        return ''
    if 'typespec' not in vars[a]:
        show(vars[a])
        outmess(f'var2fixfortran: No typespec for argument "{a}".\n')
        return ''
    vardef = vars[a]['typespec']
    if vardef == 'type' and 'typename' in vars[a]:
        vardef = f"{vardef}({vars[a]['typename']})"
    selector = {}
    lk = ''
    if 'kindselector' in vars[a]:
        selector = vars[a]['kindselector']
        lk = 'kind'
    elif 'charselector' in vars[a]:
        selector = vars[a]['charselector']
        lk = 'len'
    if '*' in selector:
        if f90mode:
            if selector['*'] in ['*', ':', '(*)']:
                vardef = f'{vardef}(len=*)'
            else:
                vardef = f"{vardef}({lk}={selector['*']})"
        elif selector['*'] in ['*', ':']:
            vardef = f"{vardef}*({selector['*']})"
        else:
            vardef = f"{vardef}*{selector['*']}"
    elif 'len' in selector:
        vardef = f"{vardef}(len={selector['len']}"
        if 'kind' in selector:
            vardef = f"{vardef},kind={selector['kind']})"
        else:
            vardef = f'{vardef})'
    elif 'kind' in selector:
        vardef = f"{vardef}(kind={selector['kind']})"

    vardef = f'{vardef} {fa}'
    if 'dimension' in vars[a]:
        vardef = f"{vardef}({','.join(vars[a]['dimension'])})"
    return vardef

def useiso_c_binding(rout):
    useisoc = False
    for key, value in rout['vars'].items():
        kind_value = value.get('kindselector', {}).get('kind')
        if kind_value in isoc_kindmap:
            return True
    return useisoc

def createfuncwrapper(rout, signature=0):
    assert isfunction(rout)

    extra_args = []
    vars = rout['vars']
    for a in rout['args']:
        v = rout['vars'][a]
        for i, d in enumerate(v.get('dimension', [])):
            if d == ':':
                dn = f'f2py_{a}_d{i}'
                dv = {'typespec': 'integer', 'intent': ['hide']}
                dv['='] = f'shape({a}, {i})'
                extra_args.append(dn)
                vars[dn] = dv
                v['dimension'][i] = dn
    rout['args'].extend(extra_args)
    need_interface = bool(extra_args)

    ret = ['']

    def add(line, ret=ret):
        ret[0] = f'{ret[0]}\n      {line}'
    name = rout['name']
    fortranname = getfortranname(rout)
    f90mode = ismoduleroutine(rout)
    newname = f'{name}f2pywrap'

    if newname not in vars:
        vars[newname] = vars[name]
        args = [newname] + rout['args'][1:]
    else:
        args = [newname] + rout['args']

    l_tmpl = var2fixfortran(vars, name, '@@@NAME@@@', f90mode)
    if l_tmpl[:13] == 'character*(*)':
        if f90mode:
            l_tmpl = 'character(len=10)' + l_tmpl[13:]
        else:
            l_tmpl = 'character*10' + l_tmpl[13:]
        charselect = vars[name]['charselector']
        if charselect.get('*', '') == '(*)':
            charselect['*'] = '10'

    l1 = l_tmpl.replace('@@@NAME@@@', newname)
    rl = None

    useisoc = useiso_c_binding(rout)
    sargs = ', '.join(args)
    if f90mode:
        # gh-23598 fix warning
        # Essentially, this gets called again with modules where the name of the
        # function is added to the arguments, which is not required, and removed
        sargs = sargs.replace(f"{name}, ", '')
        args = [arg for arg in args if arg != name]
        rout['args'] = args
        add(f"subroutine f2pywrap_{rout['modulename']}_{name} ({sargs})")
        if not signature:
            add(f"use {rout['modulename']}, only : {fortranname}")
        if useisoc:
            add('use iso_c_binding')
    else:
        add(f'subroutine f2pywrap{name} ({sargs})')
        if useisoc:
            add('use iso_c_binding')
        if not need_interface:
            add(f'external {fortranname}')
            rl = l_tmpl.replace('@@@NAME@@@', '') + ' ' + fortranname

    if need_interface:
        for line in rout['saved_interface'].split('\n'):
            if line.lstrip().startswith('use ') and '__user__' not in line:
                add(line)

    args = args[1:]
    dumped_args = []
    for a in args:
        if isexternal(vars[a]):
            add(f'external {a}')
            dumped_args.append(a)
    for a in args:
        if a in dumped_args:
            continue
        if isscalar(vars[a]):
            add(var2fixfortran(vars, a, f90mode=f90mode))
            dumped_args.append(a)
    for a in args:
        if a in dumped_args:
            continue
        if isintent_in(vars[a]):
            add(var2fixfortran(vars, a, f90mode=f90mode))
            dumped_args.append(a)
    for a in args:
        if a in dumped_args:
            continue
        add(var2fixfortran(vars, a, f90mode=f90mode))

    add(l1)
    if rl is not None:
        add(rl)

    if need_interface:
        if f90mode:
            # f90 module already defines needed interface
            pass
        else:
            add('interface')
            add(rout['saved_interface'].lstrip())
            add('end interface')

    sargs = ', '.join([a for a in args if a not in extra_args])

    if not signature:
        if islogicalfunction(rout):
            add(f'{newname} = .not.(.not.{fortranname}({sargs}))')
        else:
            add(f'{newname} = {fortranname}({sargs})')
    if f90mode:
        add(f"end subroutine f2pywrap_{rout['modulename']}_{name}")
    else:
        add('end')
    return ret[0]


def createsubrwrapper(rout, signature=0):
    assert issubroutine(rout)

    extra_args = []
    vars = rout['vars']
    for a in rout['args']:
        v = rout['vars'][a]
        for i, d in enumerate(v.get('dimension', [])):
            if d == ':':
                dn = f'f2py_{a}_d{i}'
                dv = {'typespec': 'integer', 'intent': ['hide']}
                dv['='] = f'shape({a}, {i})'
                extra_args.append(dn)
                vars[dn] = dv
                v['dimension'][i] = dn
    rout['args'].extend(extra_args)
    need_interface = bool(extra_args)

    ret = ['']

    def add(line, ret=ret):
        ret[0] = f'{ret[0]}\n      {line}'
    name = rout['name']
    fortranname = getfortranname(rout)
    f90mode = ismoduleroutine(rout)

    args = rout['args']

    useisoc = useiso_c_binding(rout)
    sargs = ', '.join(args)
    if f90mode:
        add(f"subroutine f2pywrap_{rout['modulename']}_{name} ({sargs})")
        if useisoc:
            add('use iso_c_binding')
        if not signature:
            add(f"use {rout['modulename']}, only : {fortranname}")
    else:
        add(f'subroutine f2pywrap{name} ({sargs})')
        if useisoc:
            add('use iso_c_binding')
        if not need_interface:
            add(f'external {fortranname}')

    if need_interface:
        for line in rout['saved_interface'].split('\n'):
            if line.lstrip().startswith('use ') and '__user__' not in line:
                add(line)

    dumped_args = []
    for a in args:
        if isexternal(vars[a]):
            add(f'external {a}')
            dumped_args.append(a)
    for a in args:
        if a in dumped_args:
            continue
        if isscalar(vars[a]):
            add(var2fixfortran(vars, a, f90mode=f90mode))
            dumped_args.append(a)
    for a in args:
        if a in dumped_args:
            continue
        add(var2fixfortran(vars, a, f90mode=f90mode))

    if need_interface:
        if f90mode:
            # f90 module already defines needed interface
            pass
        else:
            add('interface')
            for line in rout['saved_interface'].split('\n'):
                if line.lstrip().startswith('use ') and '__user__' in line:
                    continue
                add(line)
            add('end interface')

    sargs = ', '.join([a for a in args if a not in extra_args])

    if not signature:
        add(f'call {fortranname}({sargs})')
    if f90mode:
        add(f"end subroutine f2pywrap_{rout['modulename']}_{name}")
    else:
        add('end')
    return ret[0]


def assubr(rout):
    if isfunction_wrap(rout):
        fortranname = getfortranname(rout)
        name = rout['name']
        outmess('\t\tCreating wrapper for Fortran function "%s"("%s")...\n' % (
            name, fortranname))
        rout = copy.copy(rout)
        fname = name
        rname = fname
        if 'result' in rout:
            rname = rout['result']
            rout['vars'][fname] = rout['vars'][rname]
        fvar = rout['vars'][fname]
        if not isintent_out(fvar):
            if 'intent' not in fvar:
                fvar['intent'] = []
            fvar['intent'].append('out')
            flag = 1
            for i in fvar['intent']:
                if i.startswith('out='):
                    flag = 0
                    break
            if flag:
                fvar['intent'].append(f'out={rname}')
        rout['args'][:] = [fname] + rout['args']
        return rout, createfuncwrapper(rout)
    if issubroutine_wrap(rout):
        fortranname = getfortranname(rout)
        name = rout['name']
        outmess('\t\tCreating wrapper for Fortran subroutine "%s"("%s")...\n'
                % (name, fortranname))
        rout = copy.copy(rout)
        return rout, createsubrwrapper(rout)
    return rout, ''

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\bert\eval_squad.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
#
# This script evaluates accuracy of ONNX models for question-answering task on SQuAD data set.
# Example to evaluate raw and optimized model for CUDA in Linux:
#   pip3 install datasets evaluate optimum transformers onnxruntime-gpu
#
#   python3 eval_squad.py -m bert-large-uncased-whole-word-masking-finetuned-squad -s 384 -b 1 --use_io_binding
#
#   python3 -m onnxruntime.transformers.optimizer \
#           --input ./bert-large-uncased-whole-word-masking-finetuned-squad/model.onnx \
#           --output ./bert-large-uncased-whole-word-masking-finetuned-squad/optimized_model.onnx
#
#   python3 eval_squad.py -m bert-large-uncased-whole-word-masking-finetuned-squad -s 384 -b 1 --use_io_binding \
#           --onnx ./bert-large-uncased-whole-word-masking-finetuned-squad/optimized_model.onnx
#
#   Snippet of example output in A100:
#   {'exact': 86.65089877010406, 'f1': 92.99433524952254, 'total': 10570, 'HasAns_exact': 86.65089877010406
#    'total_time_in_seconds': 81.69239814393222, 'samples_per_second': 129.387804008115,
#    'latency_in_seconds': 0.007728703703304846, 'provider': 'CUDAExecutionProvider',
#    'pretrained_model_name': 'bert-large-uncased-whole-word-masking-finetuned-squad',
#    'batch_size': 1, 'sequence_length': 384, 'use_io_binding': True}
import argparse
import csv
import os
import time

try:
    from importlib.metadata import PackageNotFoundError, version
except ImportError:
    from importlib_metadata import PackageNotFoundError, version

from pathlib import Path
from typing import Any

from datasets import load_dataset
from evaluate import evaluator
from optimum.onnxruntime import ORTModelForQuestionAnswering
from optimum.version import __version__ as optimum_version
from packaging import version as version_check
from transformers import AutoTokenizer, pipeline

if version_check.parse(optimum_version) < version_check.parse("1.13.1"):
    raise ImportError(f"Please install optimum>=1.13.1. Current version: {optimum_version}.")

PRETRAINED_SQUAD_MODELS = [
    "bert-large-uncased-whole-word-masking-finetuned-squad",
    "deepset/roberta-base-squad2",
    "distilbert-base-cased-distilled-squad",
]


def get_package_version(package_name: str):
    try:
        return version(package_name)
    except PackageNotFoundError:
        return None


def load_onnx_model(
    model_id: str, onnx_path: str | None = None, provider="CUDAExecutionProvider", use_io_binding: bool = False
):
    """Load onnx model given pretrained model name and optional ONNX model path. If onnx_path is None,
    the default onnx model from optimum will be used.

    Args:
        model_id (str): pretrained model name or checkpoint path
        onnx_path (Optional[str], optional): path of onnx model to evaluate. Defaults to None.

    Returns:
        model: ORTModel for the onnx model
        onnx_path: the path of onnx model
    """

    if onnx_path is None:
        # Export onnx to a sub-directory named by the model id
        model = ORTModelForQuestionAnswering.from_pretrained(
            model_id, export=True, provider=provider, use_io_binding=use_io_binding
        )
        save_onnx_dir = os.path.join(".", model_id)
        model.save_pretrained(save_onnx_dir)
        onnx_path = os.path.join(save_onnx_dir, "model.onnx")
        print("Model is exported to onnx file:", onnx_path)
    else:
        model = ORTModelForQuestionAnswering.from_pretrained(
            os.path.dirname(onnx_path),
            file_name=Path(onnx_path).name,
            provider=provider,
            use_io_binding=use_io_binding,
            # provider_options={"enable_skip_layer_norm_strict_mode": True},
        )

    return model, onnx_path


def output_details(results: list[dict[str, Any]], csv_filename: str):
    """Output a CSV file with detail of each test results.

    Args:
        results (List[Dict[str, Any]]): list of JSON results.
        csv_filename (str): path of output CSV file
    """
    with open(csv_filename, mode="a", newline="", encoding="ascii") as csv_file:
        column_names = [
            "pretrained_model_name",
            "onnx_path",
            "provider",
            "disable_fused_attention",
            "batch_size",
            "sequence_length",
            "use_io_binding",
            "exact",
            "f1",
            "total",
            "HasAns_exact",
            "HasAns_f1",
            "HasAns_total",
            "best_exact",
            "best_exact_thresh",
            "best_f1",
            "best_f1_thresh",
            "total_time_in_seconds",
            "samples_per_second",
            "latency_in_seconds",
        ]

        csv_writer = csv.DictWriter(csv_file, fieldnames=column_names)
        csv_writer.writeheader()
        for result in results:
            csv_writer.writerow(result)

        csv_file.flush()

    print(f"Detail results are saved to csv file: {csv_filename}")


def output_summary(results: list[dict[str, Any]], csv_filename: str, metric_name: str):
    """Output a CSV file with summary of a metric on combinations of batch_size and sequence_length.

    Args:
        results (List[Dict[str, Any]]): list of JSON results.
        csv_filename (str): path of output CSV file
        metric_name (str): the metric to summarize
    """
    with open(csv_filename, mode="a", newline="", encoding="ascii") as csv_file:
        header_names = [
            "pretrained_model_name",
            "onnx_path",
            "provider",
            "disable_fused_attention",
            "use_io_binding",
        ]

        model_list = list({result["onnx_path"] for result in results})
        model_list.sort()

        batch_sizes = list({result["batch_size"] for result in results})
        batch_sizes.sort()

        sequence_lengths = list({result["sequence_length"] for result in results})
        sequence_lengths.sort()

        key_names = []
        for sequence_length in sequence_lengths:
            for batch_size in batch_sizes:
                key_names.append(f"b{batch_size}_s{sequence_length}")

        csv_writer = csv.DictWriter(csv_file, fieldnames=header_names + key_names)
        csv_writer.writeheader()

        for model in model_list:
            row = {}

            # Metric value for given pair of batch_size and sequence_length.
            # Assume that (onnx_path, batch_size and sequence_length) are unique so keep first occurrence only.
            values = {}
            values.update({k: "" for k in key_names})

            for result in results:
                if result["onnx_path"] == model and result[metric_name]:
                    headers = {k: v for k, v in result.items() if k in header_names}
                    if not row:
                        row.update(headers)

                    batch_size = result["batch_size"]
                    sequence_length = result["sequence_length"]
                    key = f"b{batch_size}_s{sequence_length}"

                    if key in key_names:
                        values[key] = result[metric_name]

            if row:
                for key in key_names:
                    row[key] = values.get(key, "")
                csv_writer.writerow(row)

        csv_file.flush()

    print(f"Summary results for {metric_name} are saved to csv file: {csv_filename}")


def main():
    args = parse_arguments()
    print(args)

    for name in ["onnxruntime-gpu", "onnxruntime", "onnx", "torch", "transformers", "optimum", "datasets", "evaluate"]:
        package_version = get_package_version(name)
        if package_version:
            print(f"{name} version", package_version)

    pretrained_model_name = args.model_name
    if args.onnx and not os.path.exists(args.onnx):
        raise RuntimeError(f"Onnx model path does not exist: {args.onnx}")

    disable_fused_attention = os.environ.get("ORT_DISABLE_FUSED_ATTENTION", "0") == "1"

    all_results = []
    tokenizer = AutoTokenizer.from_pretrained(pretrained_model_name)
    for sequence_length in args.sequence_lengths:
        tokenizer.model_max_length = sequence_length
        tokenizer.doc_stride = min(sequence_length // 2, 128)
        if args.onnx is None:
            print("Exporting onnx model. It might take a few minutes...")
        start_time = time.time()
        ort_model, onnx_path = load_onnx_model(pretrained_model_name, args.onnx, args.provider, args.use_io_binding)
        latency = time.time() - start_time
        print(f"Onnx model exported or loaded in {latency:.1f} seconds")

        print(ort_model.config)
        if sequence_length > ort_model.config.max_position_embeddings:
            raise RuntimeError("sequence length should not be larger than {ort_model.config.max_position_embeddings}")

        qa_pipeline = pipeline(
            "question-answering", model=ort_model, tokenizer=tokenizer, question_first=True, batch_size=args.batch_size
        )

        task_evaluator = evaluator("question-answering")
        print("Loading dataset...")
        start_time = time.time()
        squad_dataset = load_dataset("squad", split=f"validation[:{args.total}]" if args.total > 0 else "validation")
        latency = time.time() - start_time
        print(f"Dataset loaded in {latency:.1f} seconds")

        print("Evaluating squad_v2 with ORT. It might take a few minutes...")
        start_time = time.time()
        result = task_evaluator.compute(
            model_or_pipeline=qa_pipeline,
            data=squad_dataset,
            metric="squad_v2",
            squad_v2_format=True,
        )
        latency = time.time() - start_time
        print(f"Evaluation done in {latency:.1f} seconds")

        result["provider"] = args.provider
        result["disable_fused_attention"] = disable_fused_attention
        result["pretrained_model_name"] = pretrained_model_name
        result["onnx_path"] = onnx_path
        result["batch_size"] = args.batch_size
        result["sequence_length"] = sequence_length
        result["use_io_binding"] = args.use_io_binding
        print(result)

        all_results.append(result)

    output_details(all_results, "detail.csv")

    for metric_name in ["f1", "exact", "samples_per_second"]:
        output_summary(all_results, f"{metric_name}.csv", metric_name)


def parse_arguments(argv=None):
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "-m",
        "--model_name",
        required=False,
        type=str,
        default=PRETRAINED_SQUAD_MODELS[0],
        help=f"Checkpoint directory or pre-trained model names in the list: {PRETRAINED_SQUAD_MODELS}",
    )

    parser.add_argument(
        "-s",
        "--sequence_lengths",
        nargs="+",
        type=int,
        default=[384],
        help="Sequence lengths for onnx model inputs. It could have multiple values.",
    )

    parser.add_argument(
        "-b",
        "--batch_size",
        type=int,
        default=1,
        help="batch size for inference.",
    )

    parser.add_argument("-t", "--total", type=int, default=0, help="Total samples to test. 0 means all samples.")

    parser.add_argument(
        "--onnx",
        required=False,
        type=str,
        default=None,
        help="Optional onnx model path. If not specified, optimum will be used to export onnx model for testing.",
    )

    parser.add_argument(
        "--provider",
        required=False,
        default="CUDAExecutionProvider",
        help="Select which Execution Provider to use for runs. Default is CUDAExecutionProvider.",
    )

    parser.add_argument("--use_io_binding", required=False, action="store_true", help="Use IO Binding for GPU.")
    parser.set_defaults(use_io_binding=False)

    args = parser.parse_args(argv)

    return args


if __name__ == "__main__":
    main()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_core\tests\test_multiarray.py ===
import builtins
import collections.abc
import ctypes
import functools
import gc
import io
import itertools
import mmap
import operator
import os
import pathlib
import pickle
import re
import sys
import tempfile
import warnings
import weakref
from contextlib import contextmanager

# Need to test an object that does not fully implement math interface
from datetime import datetime, timedelta
from decimal import Decimal

import numpy._core._multiarray_tests as _multiarray_tests
import pytest
from numpy._core._rational_tests import rational

import numpy as np
from numpy._core.multiarray import _get_ndarray_c_version, dot
from numpy._core.tests._locales import CommaDecimalPointLocale
from numpy.exceptions import AxisError, ComplexWarning
from numpy.lib.recfunctions import repack_fields
from numpy.testing import (
    HAS_REFCOUNT,
    IS_64BIT,
    IS_PYPY,
    IS_PYSTON,
    IS_WASM,
    assert_,
    assert_allclose,
    assert_almost_equal,
    assert_array_almost_equal,
    assert_array_compare,
    assert_array_equal,
    assert_array_less,
    assert_equal,
    assert_raises,
    assert_raises_regex,
    assert_warns,
    break_cycles,
    check_support_sve,
    runstring,
    suppress_warnings,
    temppath,
)
from numpy.testing._private.utils import _no_tracing, requires_memory


def assert_arg_sorted(arr, arg):
    # resulting array should be sorted and arg values should be unique
    assert_equal(arr[arg], np.sort(arr))
    assert_equal(np.sort(arg), np.arange(len(arg)))


def assert_arr_partitioned(kth, k, arr_part):
    assert_equal(arr_part[k], kth)
    assert_array_compare(operator.__le__, arr_part[:k], kth)
    assert_array_compare(operator.__ge__, arr_part[k:], kth)


def _aligned_zeros(shape, dtype=float, order="C", align=None):
    """
    Allocate a new ndarray with aligned memory.

    The ndarray is guaranteed *not* aligned to twice the requested alignment.
    Eg, if align=4, guarantees it is not aligned to 8. If align=None uses
    dtype.alignment."""
    dtype = np.dtype(dtype)
    if dtype == np.dtype(object):
        # Can't do this, fall back to standard allocation (which
        # should always be sufficiently aligned)
        if align is not None:
            raise ValueError("object array alignment not supported")
        return np.zeros(shape, dtype=dtype, order=order)
    if align is None:
        align = dtype.alignment
    if not hasattr(shape, '__len__'):
        shape = (shape,)
    size = functools.reduce(operator.mul, shape) * dtype.itemsize
    buf = np.empty(size + 2 * align + 1, np.uint8)

    ptr = buf.__array_interface__['data'][0]
    offset = ptr % align
    if offset != 0:
        offset = align - offset
    if (ptr % (2 * align)) == 0:
        offset += align

    # Note: slices producing 0-size arrays do not necessarily change
    # data pointer --- so we use and allocate size+1
    buf = buf[offset:offset + size + 1][:-1]
    buf.fill(0)
    data = np.ndarray(shape, dtype, buf, order=order)
    return data


class TestFlags:
    def setup_method(self):
        self.a = np.arange(10)

    def test_writeable(self):
        mydict = locals()
        self.a.flags.writeable = False
        assert_raises(ValueError, runstring, 'self.a[0] = 3', mydict)
        self.a.flags.writeable = True
        self.a[0] = 5
        self.a[0] = 0

    def test_writeable_any_base(self):
        # Ensure that any base being writeable is sufficient to change flag;
        # this is especially interesting for arrays from an array interface.
        arr = np.arange(10)

        class subclass(np.ndarray):
            pass

        # Create subclass so base will not be collapsed, this is OK to change
        view1 = arr.view(subclass)
        view2 = view1[...]
        arr.flags.writeable = False
        view2.flags.writeable = False
        view2.flags.writeable = True  # Can be set to True again.

        arr = np.arange(10)

        class frominterface:
            def __init__(self, arr):
                self.arr = arr
                self.__array_interface__ = arr.__array_interface__

        view1 = np.asarray(frominterface)
        view2 = view1[...]
        view2.flags.writeable = False
        view2.flags.writeable = True

        view1.flags.writeable = False
        view2.flags.writeable = False
        with assert_raises(ValueError):
            # Must assume not writeable, since only base is not:
            view2.flags.writeable = True

    def test_writeable_from_readonly(self):
        # gh-9440 - make sure fromstring, from buffer on readonly buffers
        # set writeable False
        data = b'\x00' * 100
        vals = np.frombuffer(data, 'B')
        assert_raises(ValueError, vals.setflags, write=True)
        types = np.dtype([('vals', 'u1'), ('res3', 'S4')])
        values = np._core.records.fromstring(data, types)
        vals = values['vals']
        assert_raises(ValueError, vals.setflags, write=True)

    def test_writeable_from_buffer(self):
        data = bytearray(b'\x00' * 100)
        vals = np.frombuffer(data, 'B')
        assert_(vals.flags.writeable)
        vals.setflags(write=False)
        assert_(vals.flags.writeable is False)
        vals.setflags(write=True)
        assert_(vals.flags.writeable)
        types = np.dtype([('vals', 'u1'), ('res3', 'S4')])
        values = np._core.records.fromstring(data, types)
        vals = values['vals']
        assert_(vals.flags.writeable)
        vals.setflags(write=False)
        assert_(vals.flags.writeable is False)
        vals.setflags(write=True)
        assert_(vals.flags.writeable)

    @pytest.mark.skipif(IS_PYPY, reason="PyPy always copies")
    def test_writeable_pickle(self):
        import pickle
        # Small arrays will be copied without setting base.
        # See condition for using PyArray_SetBaseObject in
        # array_setstate.
        a = np.arange(1000)
        for v in range(pickle.HIGHEST_PROTOCOL):
            vals = pickle.loads(pickle.dumps(a, v))
            assert_(vals.flags.writeable)
            assert_(isinstance(vals.base, bytes))

    def test_writeable_from_c_data(self):
        # Test that the writeable flag can be changed for an array wrapping
        # low level C-data, but not owning its data.
        # Also see that this is deprecated to change from python.
        from numpy._core._multiarray_tests import get_c_wrapping_array

        arr_writeable = get_c_wrapping_array(True)
        assert not arr_writeable.flags.owndata
        assert arr_writeable.flags.writeable
        view = arr_writeable[...]

        # Toggling the writeable flag works on the view:
        view.flags.writeable = False
        assert not view.flags.writeable
        view.flags.writeable = True
        assert view.flags.writeable
        # Flag can be unset on the arr_writeable:
        arr_writeable.flags.writeable = False

        arr_readonly = get_c_wrapping_array(False)
        assert not arr_readonly.flags.owndata
        assert not arr_readonly.flags.writeable

        for arr in [arr_writeable, arr_readonly]:
            view = arr[...]
            view.flags.writeable = False  # make sure it is readonly
            arr.flags.writeable = False
            assert not arr.flags.writeable

            with assert_raises(ValueError):
                view.flags.writeable = True

            with assert_raises(ValueError):
                arr.flags.writeable = True

    def test_warnonwrite(self):
        a = np.arange(10)
        a.flags._warn_on_write = True
        with warnings.catch_warnings(record=True) as w:
            warnings.filterwarnings('always')
            a[1] = 10
            a[2] = 10
            # only warn once
            assert_(len(w) == 1)

    @pytest.mark.parametrize(["flag", "flag_value", "writeable"],
            [("writeable", True, True),
             # Delete _warn_on_write after deprecation and simplify
             # the parameterization:
             ("_warn_on_write", True, False),
             ("writeable", False, False)])
    def test_readonly_flag_protocols(self, flag, flag_value, writeable):
        a = np.arange(10)
        setattr(a.flags, flag, flag_value)

        class MyArr:
            __array_struct__ = a.__array_struct__

        assert memoryview(a).readonly is not writeable
        assert a.__array_interface__['data'][1] is not writeable
        assert np.asarray(MyArr()).flags.writeable is writeable

    def test_otherflags(self):
        assert_equal(self.a.flags.carray, True)
        assert_equal(self.a.flags['C'], True)
        assert_equal(self.a.flags.farray, False)
        assert_equal(self.a.flags.behaved, True)
        assert_equal(self.a.flags.fnc, False)
        assert_equal(self.a.flags.forc, True)
        assert_equal(self.a.flags.owndata, True)
        assert_equal(self.a.flags.writeable, True)
        assert_equal(self.a.flags.aligned, True)
        assert_equal(self.a.flags.writebackifcopy, False)
        assert_equal(self.a.flags['X'], False)
        assert_equal(self.a.flags['WRITEBACKIFCOPY'], False)

    def test_string_align(self):
        a = np.zeros(4, dtype=np.dtype('|S4'))
        assert_(a.flags.aligned)
        # not power of two are accessed byte-wise and thus considered aligned
        a = np.zeros(5, dtype=np.dtype('|S4'))
        assert_(a.flags.aligned)

    def test_void_align(self):
        a = np.zeros(4, dtype=np.dtype([("a", "i4"), ("b", "i4")]))
        assert_(a.flags.aligned)

    @pytest.mark.parametrize("row_size", [5, 1 << 16])
    @pytest.mark.parametrize("row_count", [1, 5])
    @pytest.mark.parametrize("ndmin", [0, 1, 2])
    def test_xcontiguous_load_txt(self, row_size, row_count, ndmin):
        s = io.StringIO('\n'.join(['1.0 ' * row_size] * row_count))
        a = np.loadtxt(s, ndmin=ndmin)

        assert a.flags.c_contiguous
        x = [i for i in a.shape if i != 1]
        assert a.flags.f_contiguous == (len(x) <= 1)


class TestHash:
    # see #3793
    def test_int(self):
        for st, ut, s in [(np.int8, np.uint8, 8),
                          (np.int16, np.uint16, 16),
                          (np.int32, np.uint32, 32),
                          (np.int64, np.uint64, 64)]:
            for i in range(1, s):
                assert_equal(hash(st(-2**i)), hash(-2**i),
                             err_msg="%r: -2**%d" % (st, i))
                assert_equal(hash(st(2**(i - 1))), hash(2**(i - 1)),
                             err_msg="%r: 2**%d" % (st, i - 1))
                assert_equal(hash(st(2**i - 1)), hash(2**i - 1),
                             err_msg="%r: 2**%d - 1" % (st, i))

                i = max(i - 1, 1)
                assert_equal(hash(ut(2**(i - 1))), hash(2**(i - 1)),
                             err_msg="%r: 2**%d" % (ut, i - 1))
                assert_equal(hash(ut(2**i - 1)), hash(2**i - 1),
                             err_msg="%r: 2**%d - 1" % (ut, i))


class TestAttributes:
    def setup_method(self):
        self.one = np.arange(10)
        self.two = np.arange(20).reshape(4, 5)
        self.three = np.arange(60, dtype=np.float64).reshape(2, 5, 6)

    def test_attributes(self):
        assert_equal(self.one.shape, (10,))
        assert_equal(self.two.shape, (4, 5))
        assert_equal(self.three.shape, (2, 5, 6))
        self.three.shape = (10, 3, 2)
        assert_equal(self.three.shape, (10, 3, 2))
        self.three.shape = (2, 5, 6)
        assert_equal(self.one.strides, (self.one.itemsize,))
        num = self.two.itemsize
        assert_equal(self.two.strides, (5 * num, num))
        num = self.three.itemsize
        assert_equal(self.three.strides, (30 * num, 6 * num, num))
        assert_equal(self.one.ndim, 1)
        assert_equal(self.two.ndim, 2)
        assert_equal(self.three.ndim, 3)
        num = self.two.itemsize
        assert_equal(self.two.size, 20)
        assert_equal(self.two.nbytes, 20 * num)
        assert_equal(self.two.itemsize, self.two.dtype.itemsize)
        assert_equal(self.two.base, np.arange(20))

    def test_dtypeattr(self):
        assert_equal(self.one.dtype, np.dtype(np.int_))
        assert_equal(self.three.dtype, np.dtype(np.float64))
        assert_equal(self.one.dtype.char, np.dtype(int).char)
        assert self.one.dtype.char in "lq"
        assert_equal(self.three.dtype.char, 'd')
        assert_(self.three.dtype.str[0] in '<>')
        assert_equal(self.one.dtype.str[1], 'i')
        assert_equal(self.three.dtype.str[1], 'f')

    def test_int_subclassing(self):
        # Regression test for https://github.com/numpy/numpy/pull/3526

        numpy_int = np.int_(0)

        # int_ doesn't inherit from Python int, because it's not fixed-width
        assert_(not isinstance(numpy_int, int))

    def test_stridesattr(self):
        x = self.one

        def make_array(size, offset, strides):
            return np.ndarray(size, buffer=x, dtype=int,
                              offset=offset * x.itemsize,
                              strides=strides * x.itemsize)

        assert_equal(make_array(4, 4, -1), np.array([4, 3, 2, 1]))
        assert_raises(ValueError, make_array, 4, 4, -2)
        assert_raises(ValueError, make_array, 4, 2, -1)
        assert_raises(ValueError, make_array, 8, 3, 1)
        assert_equal(make_array(8, 3, 0), np.array([3] * 8))
        # Check behavior reported in gh-2503:
        assert_raises(ValueError, make_array, (2, 3), 5, np.array([-2, -3]))
        make_array(0, 0, 10)

    def test_set_stridesattr(self):
        x = self.one

        def make_array(size, offset, strides):
            try:
                r = np.ndarray([size], dtype=int, buffer=x,
                               offset=offset * x.itemsize)
            except Exception as e:
                raise RuntimeError(e)
            r.strides = strides = strides * x.itemsize
            return r

        assert_equal(make_array(4, 4, -1), np.array([4, 3, 2, 1]))
        assert_equal(make_array(7, 3, 1), np.array([3, 4, 5, 6, 7, 8, 9]))
        assert_raises(ValueError, make_array, 4, 4, -2)
        assert_raises(ValueError, make_array, 4, 2, -1)
        assert_raises(RuntimeError, make_array, 8, 3, 1)
        # Check that the true extent of the array is used.
        # Test relies on as_strided base not exposing a buffer.
        x = np.lib.stride_tricks.as_strided(np.arange(1), (10, 10), (0, 0))

        def set_strides(arr, strides):
            arr.strides = strides

        assert_raises(ValueError, set_strides, x, (10 * x.itemsize, x.itemsize))

        # Test for offset calculations:
        x = np.lib.stride_tricks.as_strided(np.arange(10, dtype=np.int8)[-1],
                                                    shape=(10,), strides=(-1,))
        assert_raises(ValueError, set_strides, x[::-1], -1)
        a = x[::-1]
        a.strides = 1
        a[::2].strides = 2

        # test 0d
        arr_0d = np.array(0)
        arr_0d.strides = ()
        assert_raises(TypeError, set_strides, arr_0d, None)

    def test_fill(self):
        for t in "?bhilqpBHILQPfdgFDGO":
            x = np.empty((3, 2, 1), t)
            y = np.empty((3, 2, 1), t)
            x.fill(1)
            y[...] = 1
            assert_equal(x, y)

    def test_fill_max_uint64(self):
        x = np.empty((3, 2, 1), dtype=np.uint64)
        y = np.empty((3, 2, 1), dtype=np.uint64)
        value = 2**64 - 1
        y[...] = value
        x.fill(value)
        assert_array_equal(x, y)

    def test_fill_struct_array(self):
        # Filling from a scalar
        x = np.array([(0, 0.0), (1, 1.0)], dtype='i4,f8')
        x.fill(x[0])
        assert_equal(x['f1'][1], x['f1'][0])
        # Filling from a tuple that can be converted
        # to a scalar
        x = np.zeros(2, dtype=[('a', 'f8'), ('b', 'i4')])
        x.fill((3.5, -2))
        assert_array_equal(x['a'], [3.5, 3.5])
        assert_array_equal(x['b'], [-2, -2])

    def test_fill_readonly(self):
        # gh-22922
        a = np.zeros(11)
        a.setflags(write=False)
        with pytest.raises(ValueError, match=".*read-only"):
            a.fill(0)

    def test_fill_subarrays(self):
        # NOTE:
        # This is also a regression test for a crash with PYTHONMALLOC=debug

        dtype = np.dtype("2<i8, 2<i8, 2<i8")
        data = ([1, 2], [3, 4], [5, 6])

        arr = np.empty(1, dtype=dtype)
        arr.fill(data)

        assert_equal(arr, np.array(data, dtype=dtype))


class TestArrayConstruction:
    def test_array(self):
        d = np.ones(6)
        r = np.array([d, d])
        assert_equal(r, np.ones((2, 6)))

        d = np.ones(6)
        tgt = np.ones((2, 6))
        r = np.array([d, d])
        assert_equal(r, tgt)
        tgt[1] = 2
        r = np.array([d, d + 1])
        assert_equal(r, tgt)

        d = np.ones(6)
        r = np.array([[d, d]])
        assert_equal(r, np.ones((1, 2, 6)))

        d = np.ones(6)
        r = np.array([[d, d], [d, d]])
        assert_equal(r, np.ones((2, 2, 6)))

        d = np.ones((6, 6))
        r = np.array([d, d])
        assert_equal(r, np.ones((2, 6, 6)))

        d = np.ones((6, ))
        r = np.array([[d, d + 1], d + 2], dtype=object)
        assert_equal(len(r), 2)
        assert_equal(r[0], [d, d + 1])
        assert_equal(r[1], d + 2)

        tgt = np.ones((2, 3), dtype=bool)
        tgt[0, 2] = False
        tgt[1, 0:2] = False
        r = np.array([[True, True, False], [False, False, True]])
        assert_equal(r, tgt)
        r = np.array([[True, False], [True, False], [False, True]])
        assert_equal(r, tgt.T)

    def test_array_empty(self):
        assert_raises(TypeError, np.array)

    def test_0d_array_shape(self):
        assert np.ones(np.array(3)).shape == (3,)

    def test_array_copy_false(self):
        d = np.array([1, 2, 3])
        e = np.array(d, copy=False)
        d[1] = 3
        assert_array_equal(e, [1, 3, 3])
        np.array(d, copy=False, order='F')

    def test_array_copy_if_needed(self):
        d = np.array([1, 2, 3])
        e = np.array(d, copy=None)
        d[1] = 3
        assert_array_equal(e, [1, 3, 3])
        e = np.array(d, copy=None, order='F')
        d[1] = 4
        assert_array_equal(e, [1, 4, 3])
        e[2] = 7
        assert_array_equal(d, [1, 4, 7])

    def test_array_copy_true(self):
        d = np.array([[1, 2, 3], [1, 2, 3]])
        e = np.array(d, copy=True)
        d[0, 1] = 3
        e[0, 2] = -7
        assert_array_equal(e, [[1, 2, -7], [1, 2, 3]])
        assert_array_equal(d, [[1, 3, 3], [1, 2, 3]])
        e = np.array(d, copy=True, order='F')
        d[0, 1] = 5
        e[0, 2] = 7
        assert_array_equal(e, [[1, 3, 7], [1, 2, 3]])
        assert_array_equal(d, [[1, 5, 3], [1, 2, 3]])

    def test_array_copy_str(self):
        with pytest.raises(
            ValueError,
            match="strings are not allowed for 'copy' keyword. "
                  "Use True/False/None instead."
        ):
            np.array([1, 2, 3], copy="always")

    def test_array_cont(self):
        d = np.ones(10)[::2]
        assert_(np.ascontiguousarray(d).flags.c_contiguous)
        assert_(np.ascontiguousarray(d).flags.f_contiguous)
        assert_(np.asfortranarray(d).flags.c_contiguous)
        assert_(np.asfortranarray(d).flags.f_contiguous)
        d = np.ones((10, 10))[::2, ::2]
        assert_(np.ascontiguousarray(d).flags.c_contiguous)
        assert_(np.asfortranarray(d).flags.f_contiguous)

    @pytest.mark.parametrize("func",
            [np.array,
             np.asarray,
             np.asanyarray,
             np.ascontiguousarray,
             np.asfortranarray])
    def test_bad_arguments_error(self, func):
        with pytest.raises(TypeError):
            func(3, dtype="bad dtype")
        with pytest.raises(TypeError):
            func()  # missing arguments
        with pytest.raises(TypeError):
            func(1, 2, 3, 4, 5, 6, 7, 8)  # too many arguments

    @pytest.mark.parametrize("func",
            [np.array,
             np.asarray,
             np.asanyarray,
             np.ascontiguousarray,
             np.asfortranarray])
    def test_array_as_keyword(self, func):
        # This should likely be made positional only, but do not change
        # the name accidentally.
        if func is np.array:
            func(object=3)
        else:
            func(a=3)


class TestAssignment:
    def test_assignment_broadcasting(self):
        a = np.arange(6).reshape(2, 3)

        # Broadcasting the input to the output
        a[...] = np.arange(3)
        assert_equal(a, [[0, 1, 2], [0, 1, 2]])
        a[...] = np.arange(2).reshape(2, 1)
        assert_equal(a, [[0, 0, 0], [1, 1, 1]])

        # For compatibility with <= 1.5, a limited version of broadcasting
        # the output to the input.
        #
        # This behavior is inconsistent with NumPy broadcasting
        # in general, because it only uses one of the two broadcasting
        # rules (adding a new "1" dimension to the left of the shape),
        # applied to the output instead of an input. In NumPy 2.0, this kind
        # of broadcasting assignment will likely be disallowed.
        a[...] = np.arange(6)[::-1].reshape(1, 2, 3)
        assert_equal(a, [[5, 4, 3], [2, 1, 0]])
        # The other type of broadcasting would require a reduction operation.

        def assign(a, b):
            a[...] = b

        assert_raises(ValueError, assign, a, np.arange(12).reshape(2, 2, 3))

    def test_assignment_errors(self):
        # Address issue #2276
        class C:
            pass
        a = np.zeros(1)

        def assign(v):
            a[0] = v

        assert_raises((AttributeError, TypeError), assign, C())
        assert_raises(ValueError, assign, [1])

    @pytest.mark.filterwarnings(
        "ignore:.*set_string_function.*:DeprecationWarning"
    )
    def test_unicode_assignment(self):
        # gh-5049
        from numpy._core.arrayprint import set_printoptions

        @contextmanager
        def inject_str(s):
            """ replace ndarray.__str__ temporarily """
            set_printoptions(formatter={"all": lambda x: s})
            try:
                yield
            finally:
                set_printoptions()

        a1d = np.array(['test'])
        a0d = np.array('done')
        with inject_str('bad'):
            a1d[0] = a0d  # previously this would invoke __str__
        assert_equal(a1d[0], 'done')

        # this would crash for the same reason
        np.array([np.array('\xe5\xe4\xf6')])

    def test_stringlike_empty_list(self):
        # gh-8902
        u = np.array(['done'])
        b = np.array([b'done'])

        class bad_sequence:
            def __getitem__(self, _, /): pass
            def __len__(self): raise RuntimeError

        assert_raises(ValueError, operator.setitem, u, 0, [])
        assert_raises(ValueError, operator.setitem, b, 0, [])

        assert_raises(ValueError, operator.setitem, u, 0, bad_sequence())
        assert_raises(ValueError, operator.setitem, b, 0, bad_sequence())

    def test_longdouble_assignment(self):
        # only relevant if longdouble is larger than float
        # we're looking for loss of precision

        for dtype in (np.longdouble, np.clongdouble):
            # gh-8902
            tinyb = np.nextafter(np.longdouble(0), 1).astype(dtype)
            tinya = np.nextafter(np.longdouble(0), -1).astype(dtype)

            # construction
            tiny1d = np.array([tinya])
            assert_equal(tiny1d[0], tinya)

            # scalar = scalar
            tiny1d[0] = tinyb
            assert_equal(tiny1d[0], tinyb)

            # 0d = scalar
            tiny1d[0, ...] = tinya
            assert_equal(tiny1d[0], tinya)

            # 0d = 0d
            tiny1d[0, ...] = tinyb[...]
            assert_equal(tiny1d[0], tinyb)

            # scalar = 0d
            tiny1d[0] = tinyb[...]
            assert_equal(tiny1d[0], tinyb)

            arr = np.array([np.array(tinya)])
            assert_equal(arr[0], tinya)

    def test_cast_to_string(self):
        # cast to str should do "str(scalar)", not "str(scalar.item())"
        # When converting a float to a string via array assignment, we
        # want to ensure that the conversion uses str(scalar) to preserve
        # the expected precision.
        a = np.zeros(1, dtype='S20')
        a[:] = np.array(['1.12345678901234567890'], dtype='f8')
        assert_equal(a[0], b"1.1234567890123457")


class TestDtypedescr:
    def test_construction(self):
        d1 = np.dtype('i4')
        assert_equal(d1, np.dtype(np.int32))
        d2 = np.dtype('f8')
        assert_equal(d2, np.dtype(np.float64))

    def test_byteorders(self):
        assert_(np.dtype('<i4') != np.dtype('>i4'))
        assert_(np.dtype([('a', '<i4')]) != np.dtype([('a', '>i4')]))

    def test_structured_non_void(self):
        fields = [('a', '<i2'), ('b', '<i2')]
        dt_int = np.dtype(('i4', fields))
        assert_equal(str(dt_int), "(numpy.int32, [('a', '<i2'), ('b', '<i2')])")

        # gh-9821
        arr_int = np.zeros(4, dt_int)
        assert_equal(repr(arr_int),
            "array([0, 0, 0, 0], dtype=(numpy.int32, [('a', '<i2'), ('b', '<i2')]))")


class TestZeroRank:
    def setup_method(self):
        self.d = np.array(0), np.array('x', object)

    def test_ellipsis_subscript(self):
        a, b = self.d
        assert_equal(a[...], 0)
        assert_equal(b[...], 'x')
        assert_(a[...].base is a)  # `a[...] is a` in numpy <1.9.
        assert_(b[...].base is b)  # `b[...] is b` in numpy <1.9.

    def test_empty_subscript(self):
        a, b = self.d
        assert_equal(a[()], 0)
        assert_equal(b[()], 'x')
        assert_(type(a[()]) is a.dtype.type)
        assert_(type(b[()]) is str)

    def test_invalid_subscript(self):
        a, b = self.d
        assert_raises(IndexError, lambda x: x[0], a)
        assert_raises(IndexError, lambda x: x[0], b)
        assert_raises(IndexError, lambda x: x[np.array([], int)], a)
        assert_raises(IndexError, lambda x: x[np.array([], int)], b)

    def test_ellipsis_subscript_assignment(self):
        a, b = self.d
        a[...] = 42
        assert_equal(a, 42)
        b[...] = ''
        assert_equal(b.item(), '')

    def test_empty_subscript_assignment(self):
        a, b = self.d
        a[()] = 42
        assert_equal(a, 42)
        b[()] = ''
        assert_equal(b.item(), '')

    def test_invalid_subscript_assignment(self):
        a, b = self.d

        def assign(x, i, v):
            x[i] = v

        assert_raises(IndexError, assign, a, 0, 42)
        assert_raises(IndexError, assign, b, 0, '')
        assert_raises(ValueError, assign, a, (), '')

    def test_newaxis(self):
        a, b = self.d
        assert_equal(a[np.newaxis].shape, (1,))
        assert_equal(a[..., np.newaxis].shape, (1,))
        assert_equal(a[np.newaxis, ...].shape, (1,))
        assert_equal(a[..., np.newaxis].shape, (1,))
        assert_equal(a[np.newaxis, ..., np.newaxis].shape, (1, 1))
        assert_equal(a[..., np.newaxis, np.newaxis].shape, (1, 1))
        assert_equal(a[np.newaxis, np.newaxis, ...].shape, (1, 1))
        assert_equal(a[(np.newaxis,) * 10].shape, (1,) * 10)

    def test_invalid_newaxis(self):
        a, b = self.d

        def subscript(x, i):
            x[i]

        assert_raises(IndexError, subscript, a, (np.newaxis, 0))
        assert_raises(IndexError, subscript, a, (np.newaxis,) * 70)

    def test_constructor(self):
        x = np.ndarray(())
        x[()] = 5
        assert_equal(x[()], 5)
        y = np.ndarray((), buffer=x)
        y[()] = 6
        assert_equal(x[()], 6)

        # strides and shape must be the same length
        with pytest.raises(ValueError):
            np.ndarray((2,), strides=())
        with pytest.raises(ValueError):
            np.ndarray((), strides=(2,))

    def test_output(self):
        x = np.array(2)
        assert_raises(ValueError, np.add, x, [1], x)

    def test_real_imag(self):
        # contiguity checks are for gh-11245
        x = np.array(1j)
        xr = x.real
        xi = x.imag

        assert_equal(xr, np.array(0))
        assert_(type(xr) is np.ndarray)
        assert_equal(xr.flags.contiguous, True)
        assert_equal(xr.flags.f_contiguous, True)

        assert_equal(xi, np.array(1))
        assert_(type(xi) is np.ndarray)
        assert_equal(xi.flags.contiguous, True)
        assert_equal(xi.flags.f_contiguous, True)


class TestScalarIndexing:
    def setup_method(self):
        self.d = np.array([0, 1])[0]

    def test_ellipsis_subscript(self):
        a = self.d
        assert_equal(a[...], 0)
        assert_equal(a[...].shape, ())

    def test_empty_subscript(self):
        a = self.d
        assert_equal(a[()], 0)
        assert_equal(a[()].shape, ())

    def test_invalid_subscript(self):
        a = self.d
        assert_raises(IndexError, lambda x: x[0], a)
        assert_raises(IndexError, lambda x: x[np.array([], int)], a)

    def test_invalid_subscript_assignment(self):
        a = self.d

        def assign(x, i, v):
            x[i] = v

        assert_raises(TypeError, assign, a, 0, 42)

    def test_newaxis(self):
        a = self.d
        assert_equal(a[np.newaxis].shape, (1,))
        assert_equal(a[..., np.newaxis].shape, (1,))
        assert_equal(a[np.newaxis, ...].shape, (1,))
        assert_equal(a[..., np.newaxis].shape, (1,))
        assert_equal(a[np.newaxis, ..., np.newaxis].shape, (1, 1))
        assert_equal(a[..., np.newaxis, np.newaxis].shape, (1, 1))
        assert_equal(a[np.newaxis, np.newaxis, ...].shape, (1, 1))
        assert_equal(a[(np.newaxis,) * 10].shape, (1,) * 10)

    def test_invalid_newaxis(self):
        a = self.d

        def subscript(x, i):
            x[i]

        assert_raises(IndexError, subscript, a, (np.newaxis, 0))
        assert_raises(IndexError, subscript, a, (np.newaxis,) * 70)

    def test_overlapping_assignment(self):
        # With positive strides
        a = np.arange(4)
        a[:-1] = a[1:]
        assert_equal(a, [1, 2, 3, 3])

        a = np.arange(4)
        a[1:] = a[:-1]
        assert_equal(a, [0, 0, 1, 2])

        # With positive and negative strides
        a = np.arange(4)
        a[:] = a[::-1]
        assert_equal(a, [3, 2, 1, 0])

        a = np.arange(6).reshape(2, 3)
        a[::-1, :] = a[:, ::-1]
        assert_equal(a, [[5, 4, 3], [2, 1, 0]])

        a = np.arange(6).reshape(2, 3)
        a[::-1, ::-1] = a[:, ::-1]
        assert_equal(a, [[3, 4, 5], [0, 1, 2]])

        # With just one element overlapping
        a = np.arange(5)
        a[:3] = a[2:]
        assert_equal(a, [2, 3, 4, 3, 4])

        a = np.arange(5)
        a[2:] = a[:3]
        assert_equal(a, [0, 1, 0, 1, 2])

        a = np.arange(5)
        a[2::-1] = a[2:]
        assert_equal(a, [4, 3, 2, 3, 4])

        a = np.arange(5)
        a[2:] = a[2::-1]
        assert_equal(a, [0, 1, 2, 1, 0])

        a = np.arange(5)
        a[2::-1] = a[:1:-1]
        assert_equal(a, [2, 3, 4, 3, 4])

        a = np.arange(5)
        a[:1:-1] = a[2::-1]
        assert_equal(a, [0, 1, 0, 1, 2])


class TestCreation:
    """
    Test the np.array constructor
    """
    def test_from_attribute(self):
        class x:
            def __array__(self, dtype=None, copy=None):
                pass

        assert_raises(ValueError, np.array, x())

    def test_from_string(self):
        types = np.typecodes['AllInteger'] + np.typecodes['Float']
        nstr = ['123', '123']
        result = np.array([123, 123], dtype=int)
        for type in types:
            msg = f'String conversion for {type}'
            assert_equal(np.array(nstr, dtype=type), result, err_msg=msg)

    def test_void(self):
        arr = np.array([], dtype='V')
        assert arr.dtype == 'V8'  # current default
        # Same length scalars (those that go to the same void) work:
        arr = np.array([b"1234", b"1234"], dtype="V")
        assert arr.dtype == "V4"

        # Promoting different lengths will fail (pre 1.20 this worked)
        # by going via S5 and casting to V5.
        with pytest.raises(TypeError):
            np.array([b"1234", b"12345"], dtype="V")
        with pytest.raises(TypeError):
            np.array([b"12345", b"1234"], dtype="V")

        # Check the same for the casting path:
        arr = np.array([b"1234", b"1234"], dtype="O").astype("V")
        assert arr.dtype == "V4"
        with pytest.raises(TypeError):
            np.array([b"1234", b"12345"], dtype="O").astype("V")

    @pytest.mark.parametrize("idx",
            [pytest.param(Ellipsis, id="arr"), pytest.param((), id="scalar")])
    def test_structured_void_promotion(self, idx):
        arr = np.array(
            [np.array(1, dtype="i,i")[idx], np.array(2, dtype='i,i')[idx]],
            dtype="V")
        assert_array_equal(arr, np.array([(1, 1), (2, 2)], dtype="i,i"))
        # The following fails to promote the two dtypes, resulting in an error
        with pytest.raises(TypeError):
            np.array(
                [np.array(1, dtype="i,i")[idx], np.array(2, dtype='i,i,i')[idx]],
                dtype="V")

    def test_too_big_error(self):
        # 45341 is the smallest integer greater than sqrt(2**31 - 1).
        # 3037000500 is the smallest integer greater than sqrt(2**63 - 1).
        # We want to make sure that the square byte array with those dimensions
        # is too big on 32 or 64 bit systems respectively.
        if np.iinfo('intp').max == 2**31 - 1:
            shape = (46341, 46341)
        elif np.iinfo('intp').max == 2**63 - 1:
            shape = (3037000500, 3037000500)
        else:
            return
        assert_raises(ValueError, np.empty, shape, dtype=np.int8)
        assert_raises(ValueError, np.zeros, shape, dtype=np.int8)
        assert_raises(ValueError, np.ones, shape, dtype=np.int8)

    @pytest.mark.skipif(not IS_64BIT,
                        reason="malloc may not fail on 32 bit systems")
    def test_malloc_fails(self):
        # This test is guaranteed to fail due to a too large allocation
        with assert_raises(np._core._exceptions._ArrayMemoryError):
            np.empty(np.iinfo(np.intp).max, dtype=np.uint8)

    def test_zeros(self):
        types = np.typecodes['AllInteger'] + np.typecodes['AllFloat']
        for dt in types:
            d = np.zeros((13,), dtype=dt)
            assert_equal(np.count_nonzero(d), 0)
            # true for ieee floats
            assert_equal(d.sum(), 0)
            assert_(not d.any())

            d = np.zeros(2, dtype='(2,4)i4')
            assert_equal(np.count_nonzero(d), 0)
            assert_equal(d.sum(), 0)
            assert_(not d.any())

            d = np.zeros(2, dtype='4i4')
            assert_equal(np.count_nonzero(d), 0)
            assert_equal(d.sum(), 0)
            assert_(not d.any())

            d = np.zeros(2, dtype='(2,4)i4, (2,4)i4')
            assert_equal(np.count_nonzero(d), 0)

    @pytest.mark.slow
    def test_zeros_big(self):
        # test big array as they might be allocated different by the system
        types = np.typecodes['AllInteger'] + np.typecodes['AllFloat']
        for dt in types:
            d = np.zeros((30 * 1024**2,), dtype=dt)
            assert_(not d.any())
            # This test can fail on 32-bit systems due to insufficient
            # contiguous memory. Deallocating the previous array increases the
            # chance of success.
            del d

    def test_zeros_obj(self):
        # test initialization from PyLong(0)
        d = np.zeros((13,), dtype=object)
        assert_array_equal(d, [0] * 13)
        assert_equal(np.count_nonzero(d), 0)

    def test_zeros_obj_obj(self):
        d = np.zeros(10, dtype=[('k', object, 2)])
        assert_array_equal(d['k'], 0)

    def test_zeros_like_like_zeros(self):
        # test zeros_like returns the same as zeros
        for c in np.typecodes['All']:
            if c == 'V':
                continue
            d = np.zeros((3, 3), dtype=c)
            assert_array_equal(np.zeros_like(d), d)
            assert_equal(np.zeros_like(d).dtype, d.dtype)
        # explicitly check some special cases
        d = np.zeros((3, 3), dtype='S5')
        assert_array_equal(np.zeros_like(d), d)
        assert_equal(np.zeros_like(d).dtype, d.dtype)
        d = np.zeros((3, 3), dtype='U5')
        assert_array_equal(np.zeros_like(d), d)
        assert_equal(np.zeros_like(d).dtype, d.dtype)

        d = np.zeros((3, 3), dtype='<i4')
        assert_array_equal(np.zeros_like(d), d)
        assert_equal(np.zeros_like(d).dtype, d.dtype)
        d = np.zeros((3, 3), dtype='>i4')
        assert_array_equal(np.zeros_like(d), d)
        assert_equal(np.zeros_like(d).dtype, d.dtype)

        d = np.zeros((3, 3), dtype='<M8[s]')
        assert_array_equal(np.zeros_like(d), d)
        assert_equal(np.zeros_like(d).dtype, d.dtype)
        d = np.zeros((3, 3), dtype='>M8[s]')
        assert_array_equal(np.zeros_like(d), d)
        assert_equal(np.zeros_like(d).dtype, d.dtype)

        d = np.zeros((3, 3), dtype='f4,f4')
        assert_array_equal(np.zeros_like(d), d)
        assert_equal(np.zeros_like(d).dtype, d.dtype)

    def test_empty_unicode(self):
        # don't throw decode errors on garbage memory
        for i in range(5, 100, 5):
            d = np.empty(i, dtype='U')
            str(d)

    def test_sequence_non_homogeneous(self):
        assert_equal(np.array([4, 2**80]).dtype, object)
        assert_equal(np.array([4, 2**80, 4]).dtype, object)
        assert_equal(np.array([2**80, 4]).dtype, object)
        assert_equal(np.array([2**80] * 3).dtype, object)
        assert_equal(np.array([[1, 1], [1j, 1j]]).dtype, complex)
        assert_equal(np.array([[1j, 1j], [1, 1]]).dtype, complex)
        assert_equal(np.array([[1, 1, 1], [1, 1j, 1.], [1, 1, 1]]).dtype, complex)

    def test_non_sequence_sequence(self):
        """Should not segfault.

        Class Fail breaks the sequence protocol for new style classes, i.e.,
        those derived from object. Class Map is a mapping type indicated by
        raising a ValueError. At some point we may raise a warning instead
        of an error in the Fail case.

        """
        class Fail:
            def __len__(self):
                return 1

            def __getitem__(self, index):
                raise ValueError

        class Map:
            def __len__(self):
                return 1

            def __getitem__(self, index):
                raise KeyError

        a = np.array([Map()])
        assert_(a.shape == (1,))
        assert_(a.dtype == np.dtype(object))
        assert_raises(ValueError, np.array, [Fail()])

    def test_no_len_object_type(self):
        # gh-5100, want object array from iterable object without len()
        class Point2:
            def __init__(self):
                pass

            def __getitem__(self, ind):
                if ind in [0, 1]:
                    return ind
                else:
                    raise IndexError
        d = np.array([Point2(), Point2(), Point2()])
        assert_equal(d.dtype, np.dtype(object))

    def test_false_len_sequence(self):
        # gh-7264, segfault for this example
        class C:
            def __getitem__(self, i):
                raise IndexError

            def __len__(self):
                return 42

        a = np.array(C())  # segfault?
        assert_equal(len(a), 0)

    def test_false_len_iterable(self):
        # Special case where a bad __getitem__ makes us fall back on __iter__:
        class C:
            def __getitem__(self, x):
                raise Exception

            def __iter__(self):
                return iter(())

            def __len__(self):
                return 2

        a = np.empty(2)
        with assert_raises(ValueError):
            a[:] = C()  # Segfault!

        np.array(C()) == list(C())

    def test_failed_len_sequence(self):
        # gh-7393
        class A:
            def __init__(self, data):
                self._data = data

            def __getitem__(self, item):
                return type(self)(self._data[item])

            def __len__(self):
                return len(self._data)

        # len(d) should give 3, but len(d[0]) will fail
        d = A([1, 2, 3])
        assert_equal(len(np.array(d)), 3)

    def test_array_too_big(self):
        # Test that array creation succeeds for arrays addressable by intp
        # on the byte level and fails for too large arrays.
        buf = np.zeros(100)

        max_bytes = np.iinfo(np.intp).max
        for dtype in ["intp", "S20", "b"]:
            dtype = np.dtype(dtype)
            itemsize = dtype.itemsize

            np.ndarray(buffer=buf, strides=(0,),
                       shape=(max_bytes // itemsize,), dtype=dtype)
            assert_raises(ValueError, np.ndarray, buffer=buf, strides=(0,),
                          shape=(max_bytes // itemsize + 1,), dtype=dtype)

    def _ragged_creation(self, seq):
        # without dtype=object, the ragged object raises
        with pytest.raises(ValueError, match=".*detected shape was"):
            a = np.array(seq)

        return np.array(seq, dtype=object)

    def test_ragged_ndim_object(self):
        # Lists of mismatching depths are treated as object arrays
        a = self._ragged_creation([[1], 2, 3])
        assert_equal(a.shape, (3,))
        assert_equal(a.dtype, object)

        a = self._ragged_creation([1, [2], 3])
        assert_equal(a.shape, (3,))
        assert_equal(a.dtype, object)

        a = self._ragged_creation([1, 2, [3]])
        assert_equal(a.shape, (3,))
        assert_equal(a.dtype, object)

    def test_ragged_shape_object(self):
        # The ragged dimension of a list is turned into an object array
        a = self._ragged_creation([[1, 1], [2], [3]])
        assert_equal(a.shape, (3,))
        assert_equal(a.dtype, object)

        a = self._ragged_creation([[1], [2, 2], [3]])
        assert_equal(a.shape, (3,))
        assert_equal(a.dtype, object)

        a = self._ragged_creation([[1], [2], [3, 3]])
        assert a.shape == (3,)
        assert a.dtype == object

    def test_array_of_ragged_array(self):
        outer = np.array([None, None])
        outer[0] = outer[1] = np.array([1, 2, 3])
        assert np.array(outer).shape == (2,)
        assert np.array([outer]).shape == (1, 2)

        outer_ragged = np.array([None, None])
        outer_ragged[0] = np.array([1, 2, 3])
        outer_ragged[1] = np.array([1, 2, 3, 4])
        # should both of these emit deprecation warnings?
        assert np.array(outer_ragged).shape == (2,)
        assert np.array([outer_ragged]).shape == (1, 2,)

    def test_deep_nonragged_object(self):
        # None of these should raise, even though they are missing dtype=object
        a = np.array([[[Decimal(1)]]])
        a = np.array([1, Decimal(1)])
        a = np.array([[1], [Decimal(1)]])

    @pytest.mark.parametrize("dtype", [object, "O,O", "O,(3,)O", "(2,3)O"])
    @pytest.mark.parametrize("function", [
            np.ndarray, np.empty,
            lambda shape, dtype: np.empty_like(np.empty(shape, dtype=dtype))])
    def test_object_initialized_to_None(self, function, dtype):
        # NumPy has support for object fields to be NULL (meaning None)
        # but generally, we should always fill with the proper None, and
        # downstream may rely on that.  (For fully initialized arrays!)
        arr = function(3, dtype=dtype)
        # We expect a fill value of None, which is not NULL:
        expected = np.array(None).tobytes()
        expected = expected * (arr.nbytes // len(expected))
        assert arr.tobytes() == expected

    @pytest.mark.parametrize("func", [
        np.array, np.asarray, np.asanyarray, np.ascontiguousarray,
        np.asfortranarray])
    def test_creation_from_dtypemeta(self, func):
        dtype = np.dtype('i')
        arr1 = func([1, 2, 3], dtype=dtype)
        arr2 = func([1, 2, 3], dtype=type(dtype))
        assert_array_equal(arr1, arr2)
        assert arr2.dtype == dtype


class TestStructured:
    def test_subarray_field_access(self):
        a = np.zeros((3, 5), dtype=[('a', ('i4', (2, 2)))])
        a['a'] = np.arange(60).reshape(3, 5, 2, 2)

        # Since the subarray is always in C-order, a transpose
        # does not swap the subarray:
        assert_array_equal(a.T['a'], a['a'].transpose(1, 0, 2, 3))

        # In Fortran order, the subarray gets appended
        # like in all other cases, not prepended as a special case
        b = a.copy(order='F')
        assert_equal(a['a'].shape, b['a'].shape)
        assert_equal(a.T['a'].shape, a.T.copy()['a'].shape)

    def test_subarray_comparison(self):
        # Check that comparisons between record arrays with
        # multi-dimensional field types work properly
        a = np.rec.fromrecords(
            [([1, 2, 3], 'a', [[1, 2], [3, 4]]), ([3, 3, 3], 'b', [[0, 0], [0, 0]])],
            dtype=[('a', ('f4', 3)), ('b', object), ('c', ('i4', (2, 2)))])
        b = a.copy()
        assert_equal(a == b, [True, True])
        assert_equal(a != b, [False, False])
        b[1].b = 'c'
        assert_equal(a == b, [True, False])
        assert_equal(a != b, [False, True])
        for i in range(3):
            b[0].a = a[0].a
            b[0].a[i] = 5
            assert_equal(a == b, [False, False])
            assert_equal(a != b, [True, True])
        for i in range(2):
            for j in range(2):
                b = a.copy()
                b[0].c[i, j] = 10
                assert_equal(a == b, [False, True])
                assert_equal(a != b, [True, False])

        # Check that broadcasting with a subarray works, including cases that
        # require promotion to work:
        a = np.array([[(0,)], [(1,)]], dtype=[('a', 'f8')])
        b = np.array([(0,), (0,), (1,)], dtype=[('a', 'f8')])
        assert_equal(a == b, [[True, True, False], [False, False, True]])
        assert_equal(b == a, [[True, True, False], [False, False, True]])
        a = np.array([[(0,)], [(1,)]], dtype=[('a', 'f8', (1,))])
        b = np.array([(0,), (0,), (1,)], dtype=[('a', 'f8', (1,))])
        assert_equal(a == b, [[True, True, False], [False, False, True]])
        assert_equal(b == a, [[True, True, False], [False, False, True]])
        a = np.array([[([0, 0],)], [([1, 1],)]], dtype=[('a', 'f8', (2,))])
        b = np.array([([0, 0],), ([0, 1],), ([1, 1],)], dtype=[('a', 'f8', (2,))])
        assert_equal(a == b, [[True, False, False], [False, False, True]])
        assert_equal(b == a, [[True, False, False], [False, False, True]])

        # Check that broadcasting Fortran-style arrays with a subarray work
        a = np.array([[([0, 0],)], [([1, 1],)]], dtype=[('a', 'f8', (2,))], order='F')
        b = np.array([([0, 0],), ([0, 1],), ([1, 1],)], dtype=[('a', 'f8', (2,))])
        assert_equal(a == b, [[True, False, False], [False, False, True]])
        assert_equal(b == a, [[True, False, False], [False, False, True]])

        # Check that incompatible sub-array shapes don't result to broadcasting
        x = np.zeros((1,), dtype=[('a', ('f4', (1, 2))), ('b', 'i1')])
        y = np.zeros((1,), dtype=[('a', ('f4', (2,))), ('b', 'i1')])
        # The main importance is that it does not return True:
        with pytest.raises(TypeError):
            x == y

        x = np.zeros((1,), dtype=[('a', ('f4', (2, 1))), ('b', 'i1')])
        y = np.zeros((1,), dtype=[('a', ('f4', (2,))), ('b', 'i1')])
        # The main importance is that it does not return True:
        with pytest.raises(TypeError):
            x == y

    def test_empty_structured_array_comparison(self):
        # Check that comparison works on empty arrays with nontrivially
        # shaped fields
        a = np.zeros(0, [('a', '<f8', (1, 1))])
        assert_equal(a, a)
        a = np.zeros(0, [('a', '<f8', (1,))])
        assert_equal(a, a)
        a = np.zeros((0, 0), [('a', '<f8', (1, 1))])
        assert_equal(a, a)
        a = np.zeros((1, 0, 1), [('a', '<f8', (1, 1))])
        assert_equal(a, a)

    @pytest.mark.parametrize("op", [operator.eq, operator.ne])
    def test_structured_array_comparison_bad_broadcasts(self, op):
        a = np.zeros(3, dtype='i,i')
        b = np.array([], dtype="i,i")
        with pytest.raises(ValueError):
            op(a, b)

    def test_structured_comparisons_with_promotion(self):
        # Check that structured arrays can be compared so long as their
        # dtypes promote fine:
        a = np.array([(5, 42), (10, 1)], dtype=[('a', '>i8'), ('b', '<f8')])
        b = np.array([(5, 43), (10, 1)], dtype=[('a', '<i8'), ('b', '>f8')])
        assert_equal(a == b, [False, True])
        assert_equal(a != b, [True, False])

        a = np.array([(5, 42), (10, 1)], dtype=[('a', '>f8'), ('b', '<f8')])
        b = np.array([(5, 43), (10, 1)], dtype=[('a', '<i8'), ('b', '>i8')])
        assert_equal(a == b, [False, True])
        assert_equal(a != b, [True, False])

        # Including with embedded subarray dtype (although subarray comparison
        # itself may still be a bit weird and compare the raw data)
        a = np.array([(5, 42), (10, 1)], dtype=[('a', '10>f8'), ('b', '5<f8')])
        b = np.array([(5, 43), (10, 1)], dtype=[('a', '10<i8'), ('b', '5>i8')])
        assert_equal(a == b, [False, True])
        assert_equal(a != b, [True, False])

    @pytest.mark.parametrize("op", [
            operator.eq, lambda x, y: operator.eq(y, x),
            operator.ne, lambda x, y: operator.ne(y, x)])
    def test_void_comparison_failures(self, op):
        # In principle, one could decide to return an array of False for some
        # if comparisons are impossible.  But right now we return TypeError
        # when "void" dtype are involved.
        x = np.zeros(3, dtype=[('a', 'i1')])
        y = np.zeros(3)
        # Cannot compare non-structured to structured:
        with pytest.raises(TypeError):
            op(x, y)

        # Added title prevents promotion, but casts are OK:
        y = np.zeros(3, dtype=[(('title', 'a'), 'i1')])
        assert np.can_cast(y.dtype, x.dtype)
        with pytest.raises(TypeError):
            op(x, y)

        x = np.zeros(3, dtype="V7")
        y = np.zeros(3, dtype="V8")
        with pytest.raises(TypeError):
            op(x, y)

    def test_casting(self):
        # Check that casting a structured array to change its byte order
        # works
        a = np.array([(1,)], dtype=[('a', '<i4')])
        assert_(np.can_cast(a.dtype, [('a', '>i4')], casting='unsafe'))
        b = a.astype([('a', '>i4')])
        a_tmp = a.byteswap()
        a_tmp = a_tmp.view(a_tmp.dtype.newbyteorder())
        assert_equal(b, a_tmp)
        assert_equal(a['a'][0], b['a'][0])

        # Check that equality comparison works on structured arrays if
        # they are 'equiv'-castable
        a = np.array([(5, 42), (10, 1)], dtype=[('a', '>i4'), ('b', '<f8')])
        b = np.array([(5, 42), (10, 1)], dtype=[('a', '<i4'), ('b', '>f8')])
        assert_(np.can_cast(a.dtype, b.dtype, casting='equiv'))
        assert_equal(a == b, [True, True])

        # Check that 'equiv' casting can change byte order
        assert_(np.can_cast(a.dtype, b.dtype, casting='equiv'))
        c = a.astype(b.dtype, casting='equiv')
        assert_equal(a == c, [True, True])

        # Check that 'safe' casting can change byte order and up-cast
        # fields
        t = [('a', '<i8'), ('b', '>f8')]
        assert_(np.can_cast(a.dtype, t, casting='safe'))
        c = a.astype(t, casting='safe')
        assert_equal((c == np.array([(5, 42), (10, 1)], dtype=t)),
                     [True, True])

        # Check that 'same_kind' casting can change byte order and
        # change field widths within a "kind"
        t = [('a', '<i4'), ('b', '>f4')]
        assert_(np.can_cast(a.dtype, t, casting='same_kind'))
        c = a.astype(t, casting='same_kind')
        assert_equal((c == np.array([(5, 42), (10, 1)], dtype=t)),
                     [True, True])

        # Check that casting fails if the casting rule should fail on
        # any of the fields
        t = [('a', '>i8'), ('b', '<f4')]
        assert_(not np.can_cast(a.dtype, t, casting='safe'))
        assert_raises(TypeError, a.astype, t, casting='safe')
        t = [('a', '>i2'), ('b', '<f8')]
        assert_(not np.can_cast(a.dtype, t, casting='equiv'))
        assert_raises(TypeError, a.astype, t, casting='equiv')
        t = [('a', '>i8'), ('b', '<i2')]
        assert_(not np.can_cast(a.dtype, t, casting='same_kind'))
        assert_raises(TypeError, a.astype, t, casting='same_kind')
        assert_(not np.can_cast(a.dtype, b.dtype, casting='no'))
        assert_raises(TypeError, a.astype, b.dtype, casting='no')

        # Check that non-'unsafe' casting can't change the set of field names
        for casting in ['no', 'safe', 'equiv', 'same_kind']:
            t = [('a', '>i4')]
            assert_(not np.can_cast(a.dtype, t, casting=casting))
            t = [('a', '>i4'), ('b', '<f8'), ('c', 'i4')]
            assert_(not np.can_cast(a.dtype, t, casting=casting))

    def test_objview(self):
        # https://github.com/numpy/numpy/issues/3286
        a = np.array([], dtype=[('a', 'f'), ('b', 'f'), ('c', 'O')])
        a[['a', 'b']]  # TypeError?

        # https://github.com/numpy/numpy/issues/3253
        dat2 = np.zeros(3, [('A', 'i'), ('B', '|O')])
        dat2[['B', 'A']]  # TypeError?

    def test_setfield(self):
        # https://github.com/numpy/numpy/issues/3126
        struct_dt = np.dtype([('elem', 'i4', 5),])
        dt = np.dtype([('field', 'i4', 10), ('struct', struct_dt)])
        x = np.zeros(1, dt)
        x[0]['field'] = np.ones(10, dtype='i4')
        x[0]['struct'] = np.ones(1, dtype=struct_dt)
        assert_equal(x[0]['field'], np.ones(10, dtype='i4'))

    def test_setfield_object(self):
        # make sure object field assignment with ndarray value
        # on void scalar mimics setitem behavior
        b = np.zeros(1, dtype=[('x', 'O')])
        # next line should work identically to b['x'][0] = np.arange(3)
        b[0]['x'] = np.arange(3)
        assert_equal(b[0]['x'], np.arange(3))

        # check that broadcasting check still works
        c = np.zeros(1, dtype=[('x', 'O', 5)])

        def testassign():
            c[0]['x'] = np.arange(3)

        assert_raises(ValueError, testassign)

    def test_zero_width_string(self):
        # Test for PR #6430 / issues #473, #4955, #2585

        dt = np.dtype([('I', int), ('S', 'S0')])

        x = np.zeros(4, dtype=dt)

        assert_equal(x['S'], [b'', b'', b'', b''])
        assert_equal(x['S'].itemsize, 0)

        x['S'] = ['a', 'b', 'c', 'd']
        assert_equal(x['S'], [b'', b'', b'', b''])
        assert_equal(x['I'], [0, 0, 0, 0])

        # Variation on test case from #4955
        x['S'][x['I'] == 0] = 'hello'
        assert_equal(x['S'], [b'', b'', b'', b''])
        assert_equal(x['I'], [0, 0, 0, 0])

        # Variation on test case from #2585
        x['S'] = 'A'
        assert_equal(x['S'], [b'', b'', b'', b''])
        assert_equal(x['I'], [0, 0, 0, 0])

        # Allow zero-width dtypes in ndarray constructor
        y = np.ndarray(4, dtype=x['S'].dtype)
        assert_equal(y.itemsize, 0)
        assert_equal(x['S'], y)

        # More tests for indexing an array with zero-width fields
        assert_equal(np.zeros(4, dtype=[('a', 'S0,S0'),
                                        ('b', 'u1')])['a'].itemsize, 0)
        assert_equal(np.empty(3, dtype='S0,S0').itemsize, 0)
        assert_equal(np.zeros(4, dtype='S0,u1')['f0'].itemsize, 0)

        xx = x['S'].reshape((2, 2))
        assert_equal(xx.itemsize, 0)
        assert_equal(xx, [[b'', b''], [b'', b'']])
        # check for no uninitialized memory due to viewing S0 array
        assert_equal(xx[:].dtype, xx.dtype)
        assert_array_equal(eval(repr(xx), {"np": np, "array": np.array}), xx)

        b = io.BytesIO()
        np.save(b, xx)

        b.seek(0)
        yy = np.load(b)
        assert_equal(yy.itemsize, 0)
        assert_equal(xx, yy)

        with temppath(suffix='.npy') as tmp:
            np.save(tmp, xx)
            yy = np.load(tmp)
            assert_equal(yy.itemsize, 0)
            assert_equal(xx, yy)

    def test_base_attr(self):
        a = np.zeros(3, dtype='i4,f4')
        b = a[0]
        assert_(b.base is a)

    def test_assignment(self):
        def testassign(arr, v):
            c = arr.copy()
            c[0] = v  # assign using setitem
            c[1:] = v  # assign using "dtype_transfer" code paths
            return c

        dt = np.dtype([('foo', 'i8'), ('bar', 'i8')])
        arr = np.ones(2, dt)
        v1 = np.array([(2, 3)], dtype=[('foo', 'i8'), ('bar', 'i8')])
        v2 = np.array([(2, 3)], dtype=[('bar', 'i8'), ('foo', 'i8')])
        v3 = np.array([(2, 3)], dtype=[('bar', 'i8'), ('baz', 'i8')])
        v4 = np.array([(2,)],  dtype=[('bar', 'i8')])
        v5 = np.array([(2, 3)], dtype=[('foo', 'f8'), ('bar', 'f8')])
        w = arr.view({'names': ['bar'], 'formats': ['i8'], 'offsets': [8]})

        ans = np.array([(2, 3), (2, 3)], dtype=dt)
        assert_equal(testassign(arr, v1), ans)
        assert_equal(testassign(arr, v2), ans)
        assert_equal(testassign(arr, v3), ans)
        assert_raises(TypeError, lambda: testassign(arr, v4))
        assert_equal(testassign(arr, v5), ans)
        w[:] = 4
        assert_equal(arr, np.array([(1, 4), (1, 4)], dtype=dt))

        # test field-reordering, assignment by position, and self-assignment
        a = np.array([(1, 2, 3)],
                     dtype=[('foo', 'i8'), ('bar', 'i8'), ('baz', 'f4')])
        a[['foo', 'bar']] = a[['bar', 'foo']]
        assert_equal(a[0].item(), (2, 1, 3))

        # test that this works even for 'simple_unaligned' structs
        # (ie, that PyArray_EquivTypes cares about field order too)
        a = np.array([(1, 2)], dtype=[('a', 'i4'), ('b', 'i4')])
        a[['a', 'b']] = a[['b', 'a']]
        assert_equal(a[0].item(), (2, 1))

    def test_structuredscalar_indexing(self):
        # test gh-7262
        x = np.empty(shape=1, dtype="(2,)3S,(2,)3U")
        assert_equal(x[["f0", "f1"]][0], x[0][["f0", "f1"]])
        assert_equal(x[0], x[0][()])

    def test_multiindex_titles(self):
        a = np.zeros(4, dtype=[(('a', 'b'), 'i'), ('c', 'i'), ('d', 'i')])
        assert_raises(KeyError, lambda: a[['a', 'c']])
        assert_raises(KeyError, lambda: a[['a', 'a']])
        assert_raises(ValueError, lambda: a[['b', 'b']])  # field exists, but repeated
        a[['b', 'c']]  # no exception

    def test_structured_cast_promotion_fieldorder(self):
        # gh-15494
        # dtypes with different field names are not promotable
        A = ("a", "<i8")
        B = ("b", ">i8")
        ab = np.array([(1, 2)], dtype=[A, B])
        ba = np.array([(1, 2)], dtype=[B, A])
        assert_raises(TypeError, np.concatenate, ab, ba)
        assert_raises(TypeError, np.result_type, ab.dtype, ba.dtype)
        assert_raises(TypeError, np.promote_types, ab.dtype, ba.dtype)

        # dtypes with same field names/order but different memory offsets
        # and byte-order are promotable to packed nbo.
        assert_equal(np.promote_types(ab.dtype, ba[['a', 'b']].dtype),
                     repack_fields(ab.dtype.newbyteorder('N')))

        # gh-13667
        # dtypes with different fieldnames but castable field types are castable
        assert_equal(np.can_cast(ab.dtype, ba.dtype), True)
        assert_equal(ab.astype(ba.dtype).dtype, ba.dtype)
        assert_equal(np.can_cast('f8,i8', [('f0', 'f8'), ('f1', 'i8')]), True)
        assert_equal(np.can_cast('f8,i8', [('f1', 'f8'), ('f0', 'i8')]), True)
        assert_equal(np.can_cast('f8,i8', [('f1', 'i8'), ('f0', 'f8')]), False)
        assert_equal(np.can_cast('f8,i8', [('f1', 'i8'), ('f0', 'f8')],
                                 casting='unsafe'), True)

        ab[:] = ba  # make sure assignment still works

        # tests of type-promotion of corresponding fields
        dt1 = np.dtype([("", "i4")])
        dt2 = np.dtype([("", "i8")])
        assert_equal(np.promote_types(dt1, dt2), np.dtype([('f0', 'i8')]))
        assert_equal(np.promote_types(dt2, dt1), np.dtype([('f0', 'i8')]))
        assert_raises(TypeError, np.promote_types, dt1, np.dtype([("", "V3")]))
        assert_equal(np.promote_types('i4,f8', 'i8,f4'),
                     np.dtype([('f0', 'i8'), ('f1', 'f8')]))
        # test nested case
        dt1nest = np.dtype([("", dt1)])
        dt2nest = np.dtype([("", dt2)])
        assert_equal(np.promote_types(dt1nest, dt2nest),
                     np.dtype([('f0', np.dtype([('f0', 'i8')]))]))

        # note that offsets are lost when promoting:
        dt = np.dtype({'names': ['x'], 'formats': ['i4'], 'offsets': [8]})
        a = np.ones(3, dtype=dt)
        assert_equal(np.concatenate([a, a]).dtype, np.dtype([('x', 'i4')]))

    @pytest.mark.parametrize("dtype_dict", [
            {"names": ["a", "b"], "formats": ["i4", "f"], "itemsize": 100},
            {"names": ["a", "b"], "formats": ["i4", "f"],
                 "offsets": [0, 12]}])
    @pytest.mark.parametrize("align", [True, False])
    def test_structured_promotion_packs(self, dtype_dict, align):
        # Structured dtypes are packed when promoted (we consider the packed
        # form to be "canonical"), so tere is no extra padding.
        dtype = np.dtype(dtype_dict, align=align)
        # Remove non "canonical" dtype options:
        dtype_dict.pop("itemsize", None)
        dtype_dict.pop("offsets", None)
        expected = np.dtype(dtype_dict, align=align)

        res = np.promote_types(dtype, dtype)
        assert res.itemsize == expected.itemsize
        assert res.fields == expected.fields

        # But the "expected" one, should just be returned unchanged:
        res = np.promote_types(expected, expected)
        assert res is expected

    def test_structured_asarray_is_view(self):
        # A scalar viewing an array preserves its view even when creating a
        # new array. This test documents behaviour, it may not be the best
        # desired behaviour.
        arr = np.array([1], dtype="i,i")
        scalar = arr[0]
        assert not scalar.flags.owndata  # view into the array
        assert np.asarray(scalar).base is scalar
        # But never when a dtype is passed in:
        assert np.asarray(scalar, dtype=scalar.dtype).base is None
        # A scalar which owns its data does not have this property.
        # It is not easy to create one, one method is to use pickle:
        scalar = pickle.loads(pickle.dumps(scalar))
        assert scalar.flags.owndata
        assert np.asarray(scalar).base is None

class TestBool:
    def test_test_interning(self):
        a0 = np.bool(0)
        b0 = np.bool(False)
        assert_(a0 is b0)
        a1 = np.bool(1)
        b1 = np.bool(True)
        assert_(a1 is b1)
        assert_(np.array([True])[0] is a1)
        assert_(np.array(True)[()] is a1)

    def test_sum(self):
        d = np.ones(101, dtype=bool)
        assert_equal(d.sum(), d.size)
        assert_equal(d[::2].sum(), d[::2].size)
        assert_equal(d[::-2].sum(), d[::-2].size)

        d = np.frombuffer(b'\xff\xff' * 100, dtype=bool)
        assert_equal(d.sum(), d.size)
        assert_equal(d[::2].sum(), d[::2].size)
        assert_equal(d[::-2].sum(), d[::-2].size)

    def check_count_nonzero(self, power, length):
        powers = [2 ** i for i in range(length)]
        for i in range(2**power):
            l = [(i & x) != 0 for x in powers]
            a = np.array(l, dtype=bool)
            c = builtins.sum(l)
            assert_equal(np.count_nonzero(a), c)
            av = a.view(np.uint8)
            av *= 3
            assert_equal(np.count_nonzero(a), c)
            av *= 4
            assert_equal(np.count_nonzero(a), c)
            av[av != 0] = 0xFF
            assert_equal(np.count_nonzero(a), c)

    def test_count_nonzero(self):
        # check all 12 bit combinations in a length 17 array
        # covers most cases of the 16 byte unrolled code
        self.check_count_nonzero(12, 17)

    @pytest.mark.slow
    def test_count_nonzero_all(self):
        # check all combinations in a length 17 array
        # covers all cases of the 16 byte unrolled code
        self.check_count_nonzero(17, 17)

    def test_count_nonzero_unaligned(self):
        # prevent mistakes as e.g. gh-4060
        for o in range(7):
            a = np.zeros((18,), dtype=bool)[o + 1:]
            a[:o] = True
            assert_equal(np.count_nonzero(a), builtins.sum(a.tolist()))
            a = np.ones((18,), dtype=bool)[o + 1:]
            a[:o] = False
            assert_equal(np.count_nonzero(a), builtins.sum(a.tolist()))

    def _test_cast_from_flexible(self, dtype):
        # empty string -> false
        for n in range(3):
            v = np.array(b'', (dtype, n))
            assert_equal(bool(v), False)
            assert_equal(bool(v[()]), False)
            assert_equal(v.astype(bool), False)
            assert_(isinstance(v.astype(bool), np.ndarray))
            assert_(v[()].astype(bool) is np.False_)

        # anything else -> true
        for n in range(1, 4):
            for val in [b'a', b'0', b' ']:
                v = np.array(val, (dtype, n))
                assert_equal(bool(v), True)
                assert_equal(bool(v[()]), True)
                assert_equal(v.astype(bool), True)
                assert_(isinstance(v.astype(bool), np.ndarray))
                assert_(v[()].astype(bool) is np.True_)

    def test_cast_from_void(self):
        self._test_cast_from_flexible(np.void)

    @pytest.mark.xfail(reason="See gh-9847")
    def test_cast_from_unicode(self):
        self._test_cast_from_flexible(np.str_)

    @pytest.mark.xfail(reason="See gh-9847")
    def test_cast_from_bytes(self):
        self._test_cast_from_flexible(np.bytes_)


class TestZeroSizeFlexible:
    @staticmethod
    def _zeros(shape, dtype=str):
        dtype = np.dtype(dtype)
        if dtype == np.void:
            return np.zeros(shape, dtype=(dtype, 0))

        # not constructable directly
        dtype = np.dtype([('x', dtype, 0)])
        return np.zeros(shape, dtype=dtype)['x']

    def test_create(self):
        zs = self._zeros(10, bytes)
        assert_equal(zs.itemsize, 0)
        zs = self._zeros(10, np.void)
        assert_equal(zs.itemsize, 0)
        zs = self._zeros(10, str)
        assert_equal(zs.itemsize, 0)

    def _test_sort_partition(self, name, kinds, **kwargs):
        # Previously, these would all hang
        for dt in [bytes, np.void, str]:
            zs = self._zeros(10, dt)
            sort_method = getattr(zs, name)
            sort_func = getattr(np, name)
            for kind in kinds:
                sort_method(kind=kind, **kwargs)
                sort_func(zs, kind=kind, **kwargs)

    def test_sort(self):
        self._test_sort_partition('sort', kinds='qhs')

    def test_argsort(self):
        self._test_sort_partition('argsort', kinds='qhs')

    def test_partition(self):
        self._test_sort_partition('partition', kinds=['introselect'], kth=2)

    def test_argpartition(self):
        self._test_sort_partition('argpartition', kinds=['introselect'], kth=2)

    def test_resize(self):
        # previously an error
        for dt in [bytes, np.void, str]:
            zs = self._zeros(10, dt)
            zs.resize(25)
            zs.resize((10, 10))

    def test_view(self):
        for dt in [bytes, np.void, str]:
            zs = self._zeros(10, dt)

            # viewing as itself should be allowed
            assert_equal(zs.view(dt).dtype, np.dtype(dt))

            # viewing as any non-empty type gives an empty result
            assert_equal(zs.view((dt, 1)).shape, (0,))

    def test_dumps(self):
        zs = self._zeros(10, int)
        assert_equal(zs, pickle.loads(zs.dumps()))

    def test_pickle(self):
        for proto in range(2, pickle.HIGHEST_PROTOCOL + 1):
            for dt in [bytes, np.void, str]:
                zs = self._zeros(10, dt)
                p = pickle.dumps(zs, protocol=proto)
                zs2 = pickle.loads(p)

                assert_equal(zs.dtype, zs2.dtype)

    def test_pickle_empty(self):
        """Checking if an empty array pickled and un-pickled will not cause a
        segmentation fault"""
        arr = np.array([]).reshape(999999, 0)
        pk_dmp = pickle.dumps(arr)
        pk_load = pickle.loads(pk_dmp)

        assert pk_load.size == 0

    @pytest.mark.skipif(pickle.HIGHEST_PROTOCOL < 5,
                        reason="requires pickle protocol 5")
    def test_pickle_with_buffercallback(self):
        array = np.arange(10)
        buffers = []
        bytes_string = pickle.dumps(array, buffer_callback=buffers.append,
                                    protocol=5)
        array_from_buffer = pickle.loads(bytes_string, buffers=buffers)
        # when using pickle protocol 5 with buffer callbacks,
        # array_from_buffer is reconstructed from a buffer holding a view
        # to the initial array's data, so modifying an element in array
        # should modify it in array_from_buffer too.
        array[0] = -1
        assert array_from_buffer[0] == -1, array_from_buffer[0]


class TestMethods:

    sort_kinds = ['quicksort', 'heapsort', 'stable']

    def test_all_where(self):
        a = np.array([[True, False, True],
                      [False, False, False],
                      [True, True, True]])
        wh_full = np.array([[True, False, True],
                            [False, False, False],
                            [True, False, True]])
        wh_lower = np.array([[False],
                             [False],
                             [True]])
        for _ax in [0, None]:
            assert_equal(a.all(axis=_ax, where=wh_lower),
                        np.all(a[wh_lower[:, 0], :], axis=_ax))
            assert_equal(np.all(a, axis=_ax, where=wh_lower),
                         a[wh_lower[:, 0], :].all(axis=_ax))

        assert_equal(a.all(where=wh_full), True)
        assert_equal(np.all(a, where=wh_full), True)
        assert_equal(a.all(where=False), True)
        assert_equal(np.all(a, where=False), True)

    def test_any_where(self):
        a = np.array([[True, False, True],
                      [False, False, False],
                      [True, True, True]])
        wh_full = np.array([[False, True, False],
                            [True, True, True],
                            [False, False, False]])
        wh_middle = np.array([[False],
                              [True],
                              [False]])
        for _ax in [0, None]:
            assert_equal(a.any(axis=_ax, where=wh_middle),
                         np.any(a[wh_middle[:, 0], :], axis=_ax))
            assert_equal(np.any(a, axis=_ax, where=wh_middle),
                         a[wh_middle[:, 0], :].any(axis=_ax))
        assert_equal(a.any(where=wh_full), False)
        assert_equal(np.any(a, where=wh_full), False)
        assert_equal(a.any(where=False), False)
        assert_equal(np.any(a, where=False), False)

    @pytest.mark.parametrize("dtype",
            ["i8", "U10", "object", "datetime64[ms]"])
    def test_any_and_all_result_dtype(self, dtype):
        arr = np.ones(3, dtype=dtype)
        assert arr.any().dtype == np.bool
        assert arr.all().dtype == np.bool

    def test_any_and_all_object_dtype(self):
        # (seberg) Not sure we should even allow dtype here, but it is.
        arr = np.ones(3, dtype=object)
        # keepdims to prevent getting a scalar.
        assert arr.any(dtype=object, keepdims=True).dtype == object
        assert arr.all(dtype=object, keepdims=True).dtype == object

    def test_compress(self):
        tgt = [[5, 6, 7, 8, 9]]
        arr = np.arange(10).reshape(2, 5)
        out = arr.compress([0, 1], axis=0)
        assert_equal(out, tgt)

        tgt = [[1, 3], [6, 8]]
        out = arr.compress([0, 1, 0, 1, 0], axis=1)
        assert_equal(out, tgt)

        tgt = [[1], [6]]
        arr = np.arange(10).reshape(2, 5)
        out = arr.compress([0, 1], axis=1)
        assert_equal(out, tgt)

        arr = np.arange(10).reshape(2, 5)
        out = arr.compress([0, 1])
        assert_equal(out, 1)

    def test_choose(self):
        x = 2 * np.ones((3,), dtype=int)
        y = 3 * np.ones((3,), dtype=int)
        x2 = 2 * np.ones((2, 3), dtype=int)
        y2 = 3 * np.ones((2, 3), dtype=int)
        ind = np.array([0, 0, 1])

        A = ind.choose((x, y))
        assert_equal(A, [2, 2, 3])

        A = ind.choose((x2, y2))
        assert_equal(A, [[2, 2, 3], [2, 2, 3]])

        A = ind.choose((x, y2))
        assert_equal(A, [[2, 2, 3], [2, 2, 3]])

        oned = np.ones(1)
        # gh-12031, caused SEGFAULT
        assert_raises(TypeError, oned.choose, np.void(0), [oned])

        out = np.array(0)
        ret = np.choose(np.array(1), [10, 20, 30], out=out)
        assert out is ret
        assert_equal(out[()], 20)

        # gh-6272 check overlap on out
        x = np.arange(5)
        y = np.choose([0, 0, 0], [x[:3], x[:3], x[:3]], out=x[1:4], mode='wrap')
        assert_equal(y, np.array([0, 1, 2]))

        # gh_28206 check fail when out not writeable
        x = np.arange(3)
        out = np.zeros(3)
        out.setflags(write=False)
        assert_raises(ValueError, np.choose, [0, 1, 2], [x, x, x], out=out)

    def test_prod(self):
        ba = [1, 2, 10, 11, 6, 5, 4]
        ba2 = [[1, 2, 3, 4], [5, 6, 7, 9], [10, 3, 4, 5]]

        for ctype in [np.int16, np.uint16, np.int32, np.uint32,
                      np.float32, np.float64, np.complex64, np.complex128]:
            a = np.array(ba, ctype)
            a2 = np.array(ba2, ctype)
            if ctype in ['1', 'b']:
                assert_raises(ArithmeticError, a.prod)
                assert_raises(ArithmeticError, a2.prod, axis=1)
            else:
                assert_equal(a.prod(axis=0), 26400)
                assert_array_equal(a2.prod(axis=0),
                                   np.array([50, 36, 84, 180], ctype))
                assert_array_equal(a2.prod(axis=-1),
                                   np.array([24, 1890, 600], ctype))

    @pytest.mark.parametrize('dtype', [None, object])
    def test_repeat(self, dtype):
        m = np.array([1, 2, 3, 4, 5, 6], dtype=dtype)
        m_rect = m.reshape((2, 3))

        A = m.repeat([1, 3, 2, 1, 1, 2])
        assert_equal(A, [1, 2, 2, 2, 3,
                         3, 4, 5, 6, 6])

        A = m.repeat(2)
        assert_equal(A, [1, 1, 2, 2, 3, 3,
                         4, 4, 5, 5, 6, 6])

        A = m_rect.repeat([2, 1], axis=0)
        assert_equal(A, [[1, 2, 3],
                         [1, 2, 3],
                         [4, 5, 6]])

        A = m_rect.repeat([1, 3, 2], axis=1)
        assert_equal(A, [[1, 2, 2, 2, 3, 3],
                         [4, 5, 5, 5, 6, 6]])

        A = m_rect.repeat(2, axis=0)
        assert_equal(A, [[1, 2, 3],
                         [1, 2, 3],
                         [4, 5, 6],
                         [4, 5, 6]])

        A = m_rect.repeat(2, axis=1)
        assert_equal(A, [[1, 1, 2, 2, 3, 3],
                         [4, 4, 5, 5, 6, 6]])

    def test_reshape(self):
        arr = np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9], [10, 11, 12]])

        tgt = [[1, 2, 3, 4, 5, 6], [7, 8, 9, 10, 11, 12]]
        assert_equal(arr.reshape(2, 6), tgt)

        tgt = [[1, 2, 3, 4], [5, 6, 7, 8], [9, 10, 11, 12]]
        assert_equal(arr.reshape(3, 4), tgt)

        tgt = [[1, 10, 8, 6], [4, 2, 11, 9], [7, 5, 3, 12]]
        assert_equal(arr.reshape((3, 4), order='F'), tgt)

        tgt = [[1, 4, 7, 10], [2, 5, 8, 11], [3, 6, 9, 12]]
        assert_equal(arr.T.reshape((3, 4), order='C'), tgt)

    def test_round(self):
        def check_round(arr, expected, *round_args):
            assert_equal(arr.round(*round_args), expected)
            # With output array
            out = np.zeros_like(arr)
            res = arr.round(*round_args, out=out)
            assert_equal(out, expected)
            assert out is res

        check_round(np.array([1.2, 1.5]), [1, 2])
        check_round(np.array(1.5), 2)
        check_round(np.array([12.2, 15.5]), [10, 20], -1)
        check_round(np.array([12.15, 15.51]), [12.2, 15.5], 1)
        # Complex rounding
        check_round(np.array([4.5 + 1.5j]), [4 + 2j])
        check_round(np.array([12.5 + 15.5j]), [10 + 20j], -1)

    def test_squeeze(self):
        a = np.array([[[1], [2], [3]]])
        assert_equal(a.squeeze(), [1, 2, 3])
        assert_equal(a.squeeze(axis=(0,)), [[1], [2], [3]])
        assert_raises(ValueError, a.squeeze, axis=(1,))
        assert_equal(a.squeeze(axis=(2,)), [[1, 2, 3]])

    def test_transpose(self):
        a = np.array([[1, 2], [3, 4]])
        assert_equal(a.transpose(), [[1, 3], [2, 4]])
        assert_raises(ValueError, lambda: a.transpose(0))
        assert_raises(ValueError, lambda: a.transpose(0, 0))
        assert_raises(ValueError, lambda: a.transpose(0, 1, 2))

    def test_sort(self):
        # test ordering for floats and complex containing nans. It is only
        # necessary to check the less-than comparison, so sorts that
        # only follow the insertion sort path are sufficient. We only
        # test doubles and complex doubles as the logic is the same.

        # check doubles
        msg = "Test real sort order with nans"
        a = np.array([np.nan, 1, 0])
        b = np.sort(a)
        assert_equal(b, a[::-1], msg)
        # check complex
        msg = "Test complex sort order with nans"
        a = np.zeros(9, dtype=np.complex128)
        a.real += [np.nan, np.nan, np.nan, 1, 0, 1, 1, 0, 0]
        a.imag += [np.nan, 1, 0, np.nan, np.nan, 1, 0, 1, 0]
        b = np.sort(a)
        assert_equal(b, a[::-1], msg)

        with assert_raises_regex(
            ValueError,
            "kind` and `stable` parameters can't be provided at the same time"
        ):
            np.sort(a, kind="stable", stable=True)

    # all c scalar sorts use the same code with different types
    # so it suffices to run a quick check with one type. The number
    # of sorted items must be greater than ~50 to check the actual
    # algorithm because quick and merge sort fall over to insertion
    # sort for small arrays.

    @pytest.mark.parametrize('dtype', [np.uint8, np.uint16, np.uint32, np.uint64,
                                       np.float16, np.float32, np.float64,
                                       np.longdouble])
    def test_sort_unsigned(self, dtype):
        a = np.arange(101, dtype=dtype)
        b = a[::-1].copy()
        for kind in self.sort_kinds:
            msg = f"scalar sort, kind={kind}"
            c = a.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)
            c = b.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)

    @pytest.mark.parametrize('dtype',
                             [np.int8, np.int16, np.int32, np.int64, np.float16,
                              np.float32, np.float64, np.longdouble])
    def test_sort_signed(self, dtype):
        a = np.arange(-50, 51, dtype=dtype)
        b = a[::-1].copy()
        for kind in self.sort_kinds:
            msg = f"scalar sort, kind={kind}"
            c = a.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)
            c = b.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)

    @pytest.mark.parametrize('dtype', [np.float32, np.float64, np.longdouble])
    @pytest.mark.parametrize('part', ['real', 'imag'])
    def test_sort_complex(self, part, dtype):
        # test complex sorts. These use the same code as the scalars
        # but the compare function differs.
        cdtype = {
            np.single: np.csingle,
            np.double: np.cdouble,
            np.longdouble: np.clongdouble,
        }[dtype]
        a = np.arange(-50, 51, dtype=dtype)
        b = a[::-1].copy()
        ai = (a * (1 + 1j)).astype(cdtype)
        bi = (b * (1 + 1j)).astype(cdtype)
        setattr(ai, part, 1)
        setattr(bi, part, 1)
        for kind in self.sort_kinds:
            msg = f"complex sort, {part} part == 1, kind={kind}"
            c = ai.copy()
            c.sort(kind=kind)
            assert_equal(c, ai, msg)
            c = bi.copy()
            c.sort(kind=kind)
            assert_equal(c, ai, msg)

    def test_sort_complex_byte_swapping(self):
        # test sorting of complex arrays requiring byte-swapping, gh-5441
        for endianness in '<>':
            for dt in np.typecodes['Complex']:
                arr = np.array([1 + 3.j, 2 + 2.j, 3 + 1.j], dtype=endianness + dt)
                c = arr.copy()
                c.sort()
                msg = f'byte-swapped complex sort, dtype={dt}'
                assert_equal(c, arr, msg)

    @pytest.mark.parametrize('dtype', [np.bytes_, np.str_])
    def test_sort_string(self, dtype):
        # np.array will perform the encoding to bytes for us in the bytes test
        a = np.array(['aaaaaaaa' + chr(i) for i in range(101)], dtype=dtype)
        b = a[::-1].copy()
        for kind in self.sort_kinds:
            msg = f"kind={kind}"
            c = a.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)
            c = b.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)

    def test_sort_object(self):
        # test object array sorts.
        a = np.empty((101,), dtype=object)
        a[:] = list(range(101))
        b = a[::-1]
        for kind in ['q', 'h', 'm']:
            msg = f"kind={kind}"
            c = a.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)
            c = b.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)

    @pytest.mark.parametrize("dt", [
            np.dtype([('f', float), ('i', int)]),
            np.dtype([('f', float), ('i', object)])])
    @pytest.mark.parametrize("step", [1, 2])
    def test_sort_structured(self, dt, step):
        # test record array sorts.
        a = np.array([(i, i) for i in range(101 * step)], dtype=dt)
        b = a[::-1]
        for kind in ['q', 'h', 'm']:
            msg = f"kind={kind}"
            c = a.copy()[::step]
            indx = c.argsort(kind=kind)
            c.sort(kind=kind)
            assert_equal(c, a[::step], msg)
            assert_equal(a[::step][indx], a[::step], msg)
            c = b.copy()[::step]
            indx = c.argsort(kind=kind)
            c.sort(kind=kind)
            assert_equal(c, a[step - 1::step], msg)
            assert_equal(b[::step][indx], a[step - 1::step], msg)

    @pytest.mark.parametrize('dtype', ['datetime64[D]', 'timedelta64[D]'])
    def test_sort_time(self, dtype):
        # test datetime64 and timedelta64 sorts.
        a = np.arange(0, 101, dtype=dtype)
        b = a[::-1]
        for kind in ['q', 'h', 'm']:
            msg = f"kind={kind}"
            c = a.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)
            c = b.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)

    def test_sort_axis(self):
        # check axis handling. This should be the same for all type
        # specific sorts, so we only check it for one type and one kind
        a = np.array([[3, 2], [1, 0]])
        b = np.array([[1, 0], [3, 2]])
        c = np.array([[2, 3], [0, 1]])
        d = a.copy()
        d.sort(axis=0)
        assert_equal(d, b, "test sort with axis=0")
        d = a.copy()
        d.sort(axis=1)
        assert_equal(d, c, "test sort with axis=1")
        d = a.copy()
        d.sort()
        assert_equal(d, c, "test sort with default axis")

    def test_sort_size_0(self):
        # check axis handling for multidimensional empty arrays
        a = np.array([])
        a.shape = (3, 2, 1, 0)
        for axis in range(-a.ndim, a.ndim):
            msg = f'test empty array sort with axis={axis}'
            assert_equal(np.sort(a, axis=axis), a, msg)
        msg = 'test empty array sort with axis=None'
        assert_equal(np.sort(a, axis=None), a.ravel(), msg)

    def test_sort_bad_ordering(self):
        # test generic class with bogus ordering,
        # should not segfault.
        class Boom:
            def __lt__(self, other):
                return True

        a = np.array([Boom()] * 100, dtype=object)
        for kind in self.sort_kinds:
            msg = f"kind={kind}"
            c = a.copy()
            c.sort(kind=kind)
            assert_equal(c, a, msg)

    def test_void_sort(self):
        # gh-8210 - previously segfaulted
        for i in range(4):
            rand = np.random.randint(256, size=4000, dtype=np.uint8)
            arr = rand.view('V4')
            arr[::-1].sort()

        dt = np.dtype([('val', 'i4', (1,))])
        for i in range(4):
            rand = np.random.randint(256, size=4000, dtype=np.uint8)
            arr = rand.view(dt)
            arr[::-1].sort()

    def test_sort_raises(self):
        # gh-9404
        arr = np.array([0, datetime.now(), 1], dtype=object)
        for kind in self.sort_kinds:
            assert_raises(TypeError, arr.sort, kind=kind)
        # gh-3879

        class Raiser:
            def raises_anything(*args, **kwargs):
                raise TypeError("SOMETHING ERRORED")
            __eq__ = __ne__ = __lt__ = __gt__ = __ge__ = __le__ = raises_anything
        arr = np.array([[Raiser(), n] for n in range(10)]).reshape(-1)
        np.random.shuffle(arr)
        for kind in self.sort_kinds:
            assert_raises(TypeError, arr.sort, kind=kind)

    def test_sort_degraded(self):
        # test degraded dataset would take minutes to run with normal qsort
        d = np.arange(1000000)
        do = d.copy()
        x = d
        # create a median of 3 killer where each median is the sorted second
        # last element of the quicksort partition
        while x.size > 3:
            mid = x.size // 2
            x[mid], x[-2] = x[-2], x[mid]
            x = x[:-2]

        assert_equal(np.sort(d), do)
        assert_equal(d[np.argsort(d)], do)

    def test_copy(self):
        def assert_fortran(arr):
            assert_(arr.flags.fortran)
            assert_(arr.flags.f_contiguous)
            assert_(not arr.flags.c_contiguous)

        def assert_c(arr):
            assert_(not arr.flags.fortran)
            assert_(not arr.flags.f_contiguous)
            assert_(arr.flags.c_contiguous)

        a = np.empty((2, 2), order='F')
        # Test copying a Fortran array
        assert_c(a.copy())
        assert_c(a.copy('C'))
        assert_fortran(a.copy('F'))
        assert_fortran(a.copy('A'))

        # Now test starting with a C array.
        a = np.empty((2, 2), order='C')
        assert_c(a.copy())
        assert_c(a.copy('C'))
        assert_fortran(a.copy('F'))
        assert_c(a.copy('A'))

    @pytest.mark.parametrize("dtype", ['O', np.int32, 'i,O'])
    def test__deepcopy__(self, dtype):
        # Force the entry of NULLs into array
        a = np.empty(4, dtype=dtype)
        ctypes.memset(a.ctypes.data, 0, a.nbytes)

        # Ensure no error is raised, see gh-21833
        b = a.__deepcopy__({})

        a[0] = 42
        with pytest.raises(AssertionError):
            assert_array_equal(a, b)

    def test__deepcopy__catches_failure(self):
        class MyObj:
            def __deepcopy__(self, *args, **kwargs):
                raise RuntimeError

        arr = np.array([1, MyObj(), 3], dtype='O')
        with pytest.raises(RuntimeError):
            arr.__deepcopy__({})

    def test_sort_order(self):
        # Test sorting an array with fields
        x1 = np.array([21, 32, 14])
        x2 = np.array(['my', 'first', 'name'])
        x3 = np.array([3.1, 4.5, 6.2])
        r = np.rec.fromarrays([x1, x2, x3], names='id,word,number')

        r.sort(order=['id'])
        assert_equal(r.id, np.array([14, 21, 32]))
        assert_equal(r.word, np.array(['name', 'my', 'first']))
        assert_equal(r.number, np.array([6.2, 3.1, 4.5]))

        r.sort(order=['word'])
        assert_equal(r.id, np.array([32, 21, 14]))
        assert_equal(r.word, np.array(['first', 'my', 'name']))
        assert_equal(r.number, np.array([4.5, 3.1, 6.2]))

        r.sort(order=['number'])
        assert_equal(r.id, np.array([21, 32, 14]))
        assert_equal(r.word, np.array(['my', 'first', 'name']))
        assert_equal(r.number, np.array([3.1, 4.5, 6.2]))

        assert_raises_regex(ValueError, 'duplicate',
            lambda: r.sort(order=['id', 'id']))

        if sys.byteorder == 'little':
            strtype = '>i2'
        else:
            strtype = '<i2'
        mydtype = [('name', 'U5'), ('col2', strtype)]
        r = np.array([('a', 1), ('b', 255), ('c', 3), ('d', 258)],
                     dtype=mydtype)
        r.sort(order='col2')
        assert_equal(r['col2'], [1, 3, 255, 258])
        assert_equal(r, np.array([('a', 1), ('c', 3), ('b', 255), ('d', 258)],
                                 dtype=mydtype))

    def test_argsort(self):
        # all c scalar argsorts use the same code with different types
        # so it suffices to run a quick check with one type. The number
        # of sorted items must be greater than ~50 to check the actual
        # algorithm because quick and merge sort fall over to insertion
        # sort for small arrays.

        for dtype in [np.int32, np.uint32, np.float32]:
            a = np.arange(101, dtype=dtype)
            b = a[::-1].copy()
            for kind in self.sort_kinds:
                msg = f"scalar argsort, kind={kind}, dtype={dtype}"
                assert_equal(a.copy().argsort(kind=kind), a, msg)
                assert_equal(b.copy().argsort(kind=kind), b, msg)

        # test complex argsorts. These use the same code as the scalars
        # but the compare function differs.
        ai = a * 1j + 1
        bi = b * 1j + 1
        for kind in self.sort_kinds:
            msg = f"complex argsort, kind={kind}"
            assert_equal(ai.copy().argsort(kind=kind), a, msg)
            assert_equal(bi.copy().argsort(kind=kind), b, msg)
        ai = a + 1j
        bi = b + 1j
        for kind in self.sort_kinds:
            msg = f"complex argsort, kind={kind}"
            assert_equal(ai.copy().argsort(kind=kind), a, msg)
            assert_equal(bi.copy().argsort(kind=kind), b, msg)

        # test argsort of complex arrays requiring byte-swapping, gh-5441
        for endianness in '<>':
            for dt in np.typecodes['Complex']:
                arr = np.array([1 + 3.j, 2 + 2.j, 3 + 1.j], dtype=endianness + dt)
                msg = f'byte-swapped complex argsort, dtype={dt}'
                assert_equal(arr.argsort(),
                             np.arange(len(arr), dtype=np.intp), msg)

        # test string argsorts.
        s = 'aaaaaaaa'
        a = np.array([s + chr(i) for i in range(101)])
        b = a[::-1].copy()
        r = np.arange(101)
        rr = r[::-1]
        for kind in self.sort_kinds:
            msg = f"string argsort, kind={kind}"
            assert_equal(a.copy().argsort(kind=kind), r, msg)
            assert_equal(b.copy().argsort(kind=kind), rr, msg)

        # test unicode argsorts.
        s = 'aaaaaaaa'
        a = np.array([s + chr(i) for i in range(101)], dtype=np.str_)
        b = a[::-1]
        r = np.arange(101)
        rr = r[::-1]
        for kind in self.sort_kinds:
            msg = f"unicode argsort, kind={kind}"
            assert_equal(a.copy().argsort(kind=kind), r, msg)
            assert_equal(b.copy().argsort(kind=kind), rr, msg)

        # test object array argsorts.
        a = np.empty((101,), dtype=object)
        a[:] = list(range(101))
        b = a[::-1]
        r = np.arange(101)
        rr = r[::-1]
        for kind in self.sort_kinds:
            msg = f"object argsort, kind={kind}"
            assert_equal(a.copy().argsort(kind=kind), r, msg)
            assert_equal(b.copy().argsort(kind=kind), rr, msg)

        # test structured array argsorts.
        dt = np.dtype([('f', float), ('i', int)])
        a = np.array([(i, i) for i in range(101)], dtype=dt)
        b = a[::-1]
        r = np.arange(101)
        rr = r[::-1]
        for kind in self.sort_kinds:
            msg = f"structured array argsort, kind={kind}"
            assert_equal(a.copy().argsort(kind=kind), r, msg)
            assert_equal(b.copy().argsort(kind=kind), rr, msg)

        # test datetime64 argsorts.
        a = np.arange(0, 101, dtype='datetime64[D]')
        b = a[::-1]
        r = np.arange(101)
        rr = r[::-1]
        for kind in ['q', 'h', 'm']:
            msg = f"datetime64 argsort, kind={kind}"
            assert_equal(a.copy().argsort(kind=kind), r, msg)
            assert_equal(b.copy().argsort(kind=kind), rr, msg)

        # test timedelta64 argsorts.
        a = np.arange(0, 101, dtype='timedelta64[D]')
        b = a[::-1]
        r = np.arange(101)
        rr = r[::-1]
        for kind in ['q', 'h', 'm']:
            msg = f"timedelta64 argsort, kind={kind}"
            assert_equal(a.copy().argsort(kind=kind), r, msg)
            assert_equal(b.copy().argsort(kind=kind), rr, msg)

        # check axis handling. This should be the same for all type
        # specific argsorts, so we only check it for one type and one kind
        a = np.array([[3, 2], [1, 0]])
        b = np.array([[1, 1], [0, 0]])
        c = np.array([[1, 0], [1, 0]])
        assert_equal(a.copy().argsort(axis=0), b)
        assert_equal(a.copy().argsort(axis=1), c)
        assert_equal(a.copy().argsort(), c)

        # check axis handling for multidimensional empty arrays
        a = np.array([])
        a.shape = (3, 2, 1, 0)
        for axis in range(-a.ndim, a.ndim):
            msg = f'test empty array argsort with axis={axis}'
            assert_equal(np.argsort(a, axis=axis),
                         np.zeros_like(a, dtype=np.intp), msg)
        msg = 'test empty array argsort with axis=None'
        assert_equal(np.argsort(a, axis=None),
                     np.zeros_like(a.ravel(), dtype=np.intp), msg)

        # check that stable argsorts are stable
        r = np.arange(100)
        # scalars
        a = np.zeros(100)
        assert_equal(a.argsort(kind='m'), r)
        # complex
        a = np.zeros(100, dtype=complex)
        assert_equal(a.argsort(kind='m'), r)
        # string
        a = np.array(['aaaaaaaaa' for i in range(100)])
        assert_equal(a.argsort(kind='m'), r)
        # unicode
        a = np.array(['aaaaaaaaa' for i in range(100)], dtype=np.str_)
        assert_equal(a.argsort(kind='m'), r)

        with assert_raises_regex(
            ValueError,
            "kind` and `stable` parameters can't be provided at the same time"
        ):
            np.argsort(a, kind="stable", stable=True)

    def test_sort_unicode_kind(self):
        d = np.arange(10)
        k = b'\xc3\xa4'.decode("UTF8")
        assert_raises(ValueError, d.sort, kind=k)
        assert_raises(ValueError, d.argsort, kind=k)

    @pytest.mark.parametrize('a', [
        np.array([0, 1, np.nan], dtype=np.float16),
        np.array([0, 1, np.nan], dtype=np.float32),
        np.array([0, 1, np.nan]),
    ])
    def test_searchsorted_floats(self, a):
        # test for floats arrays containing nans. Explicitly test
        # half, single, and double precision floats to verify that
        # the NaN-handling is correct.
        msg = f"Test real ({a.dtype}) searchsorted with nans, side='l'"
        b = a.searchsorted(a, side='left')
        assert_equal(b, np.arange(3), msg)
        msg = f"Test real ({a.dtype}) searchsorted with nans, side='r'"
        b = a.searchsorted(a, side='right')
        assert_equal(b, np.arange(1, 4), msg)
        # check keyword arguments
        a.searchsorted(v=1)
        x = np.array([0, 1, np.nan], dtype='float32')
        y = np.searchsorted(x, x[-1])
        assert_equal(y, 2)

    def test_searchsorted_complex(self):
        # test for complex arrays containing nans.
        # The search sorted routines use the compare functions for the
        # array type, so this checks if that is consistent with the sort
        # order.
        # check double complex
        a = np.zeros(9, dtype=np.complex128)
        a.real += [0, 0, 1, 1, 0, 1, np.nan, np.nan, np.nan]
        a.imag += [0, 1, 0, 1, np.nan, np.nan, 0, 1, np.nan]
        msg = "Test complex searchsorted with nans, side='l'"
        b = a.searchsorted(a, side='left')
        assert_equal(b, np.arange(9), msg)
        msg = "Test complex searchsorted with nans, side='r'"
        b = a.searchsorted(a, side='right')
        assert_equal(b, np.arange(1, 10), msg)
        msg = "Test searchsorted with little endian, side='l'"
        a = np.array([0, 128], dtype='<i4')
        b = a.searchsorted(np.array(128, dtype='<i4'))
        assert_equal(b, 1, msg)
        msg = "Test searchsorted with big endian, side='l'"
        a = np.array([0, 128], dtype='>i4')
        b = a.searchsorted(np.array(128, dtype='>i4'))
        assert_equal(b, 1, msg)

    def test_searchsorted_n_elements(self):
        # Check 0 elements
        a = np.ones(0)
        b = a.searchsorted([0, 1, 2], 'left')
        assert_equal(b, [0, 0, 0])
        b = a.searchsorted([0, 1, 2], 'right')
        assert_equal(b, [0, 0, 0])
        a = np.ones(1)
        # Check 1 element
        b = a.searchsorted([0, 1, 2], 'left')
        assert_equal(b, [0, 0, 1])
        b = a.searchsorted([0, 1, 2], 'right')
        assert_equal(b, [0, 1, 1])
        # Check all elements equal
        a = np.ones(2)
        b = a.searchsorted([0, 1, 2], 'left')
        assert_equal(b, [0, 0, 2])
        b = a.searchsorted([0, 1, 2], 'right')
        assert_equal(b, [0, 2, 2])

    def test_searchsorted_unaligned_array(self):
        # Test searching unaligned array
        a = np.arange(10)
        aligned = np.empty(a.itemsize * a.size + 1, 'uint8')
        unaligned = aligned[1:].view(a.dtype)
        unaligned[:] = a
        # Test searching unaligned array
        b = unaligned.searchsorted(a, 'left')
        assert_equal(b, a)
        b = unaligned.searchsorted(a, 'right')
        assert_equal(b, a + 1)
        # Test searching for unaligned keys
        b = a.searchsorted(unaligned, 'left')
        assert_equal(b, a)
        b = a.searchsorted(unaligned, 'right')
        assert_equal(b, a + 1)

    def test_searchsorted_resetting(self):
        # Test smart resetting of binsearch indices
        a = np.arange(5)
        b = a.searchsorted([6, 5, 4], 'left')
        assert_equal(b, [5, 5, 4])
        b = a.searchsorted([6, 5, 4], 'right')
        assert_equal(b, [5, 5, 5])

    def test_searchsorted_type_specific(self):
        # Test all type specific binary search functions
        types = ''.join((np.typecodes['AllInteger'], np.typecodes['AllFloat'],
                         np.typecodes['Datetime'], '?O'))
        for dt in types:
            if dt == 'M':
                dt = 'M8[D]'
            if dt == '?':
                a = np.arange(2, dtype=dt)
                out = np.arange(2)
            else:
                a = np.arange(0, 5, dtype=dt)
                out = np.arange(5)
            b = a.searchsorted(a, 'left')
            assert_equal(b, out)
            b = a.searchsorted(a, 'right')
            assert_equal(b, out + 1)
            # Test empty array, use a fresh array to get warnings in
            # valgrind if access happens.
            e = np.ndarray(shape=0, buffer=b'', dtype=dt)
            b = e.searchsorted(a, 'left')
            assert_array_equal(b, np.zeros(len(a), dtype=np.intp))
            b = a.searchsorted(e, 'left')
            assert_array_equal(b, np.zeros(0, dtype=np.intp))

    def test_searchsorted_unicode(self):
        # Test searchsorted on unicode strings.

        # 1.6.1 contained a string length miscalculation in
        # arraytypes.c.src:UNICODE_compare() which manifested as
        # incorrect/inconsistent results from searchsorted.
        a = np.array(['P:\\20x_dapi_cy3\\20x_dapi_cy3_20100185_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100186_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100187_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100189_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100190_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100191_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100192_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100193_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100194_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100195_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100196_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100197_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100198_1',
                      'P:\\20x_dapi_cy3\\20x_dapi_cy3_20100199_1'],
                     dtype=np.str_)
        ind = np.arange(len(a))
        assert_equal([a.searchsorted(v, 'left') for v in a], ind)
        assert_equal([a.searchsorted(v, 'right') for v in a], ind + 1)
        assert_equal([a.searchsorted(a[i], 'left') for i in ind], ind)
        assert_equal([a.searchsorted(a[i], 'right') for i in ind], ind + 1)

    def test_searchsorted_with_invalid_sorter(self):
        a = np.array([5, 2, 1, 3, 4])
        s = np.argsort(a)
        assert_raises(TypeError, np.searchsorted, a, 0,
                      sorter=np.array((1, (2, 3)), dtype=object))
        assert_raises(TypeError, np.searchsorted, a, 0, sorter=[1.1])
        assert_raises(ValueError, np.searchsorted, a, 0, sorter=[1, 2, 3, 4])
        assert_raises(ValueError, np.searchsorted, a, 0, sorter=[1, 2, 3, 4, 5, 6])

        # bounds check
        assert_raises(ValueError, np.searchsorted, a, 4, sorter=[0, 1, 2, 3, 5])
        assert_raises(ValueError, np.searchsorted, a, 0, sorter=[-1, 0, 1, 2, 3])
        assert_raises(ValueError, np.searchsorted, a, 0, sorter=[4, 0, -1, 2, 3])

    def test_searchsorted_with_sorter(self):
        a = np.random.rand(300)
        s = a.argsort()
        b = np.sort(a)
        k = np.linspace(0, 1, 20)
        assert_equal(b.searchsorted(k), a.searchsorted(k, sorter=s))

        a = np.array([0, 1, 2, 3, 5] * 20)
        s = a.argsort()
        k = [0, 1, 2, 3, 5]
        expected = [0, 20, 40, 60, 80]
        assert_equal(a.searchsorted(k, side='left', sorter=s), expected)
        expected = [20, 40, 60, 80, 100]
        assert_equal(a.searchsorted(k, side='right', sorter=s), expected)

        # Test searching unaligned array
        keys = np.arange(10)
        a = keys.copy()
        np.random.shuffle(s)
        s = a.argsort()
        aligned = np.empty(a.itemsize * a.size + 1, 'uint8')
        unaligned = aligned[1:].view(a.dtype)
        # Test searching unaligned array
        unaligned[:] = a
        b = unaligned.searchsorted(keys, 'left', s)
        assert_equal(b, keys)
        b = unaligned.searchsorted(keys, 'right', s)
        assert_equal(b, keys + 1)
        # Test searching for unaligned keys
        unaligned[:] = keys
        b = a.searchsorted(unaligned, 'left', s)
        assert_equal(b, keys)
        b = a.searchsorted(unaligned, 'right', s)
        assert_equal(b, keys + 1)

        # Test all type specific indirect binary search functions
        types = ''.join((np.typecodes['AllInteger'], np.typecodes['AllFloat'],
                         np.typecodes['Datetime'], '?O'))
        for dt in types:
            if dt == 'M':
                dt = 'M8[D]'
            if dt == '?':
                a = np.array([1, 0], dtype=dt)
                # We want the sorter array to be of a type that is different
                # from np.intp in all platforms, to check for #4698
                s = np.array([1, 0], dtype=np.int16)
                out = np.array([1, 0])
            else:
                a = np.array([3, 4, 1, 2, 0], dtype=dt)
                # We want the sorter array to be of a type that is different
                # from np.intp in all platforms, to check for #4698
                s = np.array([4, 2, 3, 0, 1], dtype=np.int16)
                out = np.array([3, 4, 1, 2, 0], dtype=np.intp)
            b = a.searchsorted(a, 'left', s)
            assert_equal(b, out)
            b = a.searchsorted(a, 'right', s)
            assert_equal(b, out + 1)
            # Test empty array, use a fresh array to get warnings in
            # valgrind if access happens.
            e = np.ndarray(shape=0, buffer=b'', dtype=dt)
            b = e.searchsorted(a, 'left', s[:0])
            assert_array_equal(b, np.zeros(len(a), dtype=np.intp))
            b = a.searchsorted(e, 'left', s)
            assert_array_equal(b, np.zeros(0, dtype=np.intp))

        # Test non-contiguous sorter array
        a = np.array([3, 4, 1, 2, 0])
        srt = np.empty((10,), dtype=np.intp)
        srt[1::2] = -1
        srt[::2] = [4, 2, 3, 0, 1]
        s = srt[::2]
        out = np.array([3, 4, 1, 2, 0], dtype=np.intp)
        b = a.searchsorted(a, 'left', s)
        assert_equal(b, out)
        b = a.searchsorted(a, 'right', s)
        assert_equal(b, out + 1)

    def test_searchsorted_return_type(self):
        # Functions returning indices should always return base ndarrays
        class A(np.ndarray):
            pass
        a = np.arange(5).view(A)
        b = np.arange(1, 3).view(A)
        s = np.arange(5).view(A)
        assert_(not isinstance(a.searchsorted(b, 'left'), A))
        assert_(not isinstance(a.searchsorted(b, 'right'), A))
        assert_(not isinstance(a.searchsorted(b, 'left', s), A))
        assert_(not isinstance(a.searchsorted(b, 'right', s), A))

    @pytest.mark.parametrize("dtype", np.typecodes["All"])
    def test_argpartition_out_of_range(self, dtype):
        # Test out of range values in kth raise an error, gh-5469
        d = np.arange(10).astype(dtype=dtype)
        assert_raises(ValueError, d.argpartition, 10)
        assert_raises(ValueError, d.argpartition, -11)

    @pytest.mark.parametrize("dtype", np.typecodes["All"])
    def test_partition_out_of_range(self, dtype):
        # Test out of range values in kth raise an error, gh-5469
        d = np.arange(10).astype(dtype=dtype)
        assert_raises(ValueError, d.partition, 10)
        assert_raises(ValueError, d.partition, -11)

    def test_argpartition_integer(self):
        # Test non-integer values in kth raise an error/
        d = np.arange(10)
        assert_raises(TypeError, d.argpartition, 9.)
        # Test also for generic type argpartition, which uses sorting
        # and used to not bound check kth
        d_obj = np.arange(10, dtype=object)
        assert_raises(TypeError, d_obj.argpartition, 9.)

    def test_partition_integer(self):
        # Test out of range values in kth raise an error, gh-5469
        d = np.arange(10)
        assert_raises(TypeError, d.partition, 9.)
        # Test also for generic type partition, which uses sorting
        # and used to not bound check kth
        d_obj = np.arange(10, dtype=object)
        assert_raises(TypeError, d_obj.partition, 9.)

    @pytest.mark.parametrize("kth_dtype", np.typecodes["AllInteger"])
    def test_partition_empty_array(self, kth_dtype):
        # check axis handling for multidimensional empty arrays
        kth = np.array(0, dtype=kth_dtype)[()]
        a = np.array([])
        a.shape = (3, 2, 1, 0)
        for axis in range(-a.ndim, a.ndim):
            msg = f'test empty array partition with axis={axis}'
            assert_equal(np.partition(a, kth, axis=axis), a, msg)
        msg = 'test empty array partition with axis=None'
        assert_equal(np.partition(a, kth, axis=None), a.ravel(), msg)

    @pytest.mark.parametrize("kth_dtype", np.typecodes["AllInteger"])
    def test_argpartition_empty_array(self, kth_dtype):
        # check axis handling for multidimensional empty arrays
        kth = np.array(0, dtype=kth_dtype)[()]
        a = np.array([])
        a.shape = (3, 2, 1, 0)
        for axis in range(-a.ndim, a.ndim):
            msg = f'test empty array argpartition with axis={axis}'
            assert_equal(np.partition(a, kth, axis=axis),
                         np.zeros_like(a, dtype=np.intp), msg)
        msg = 'test empty array argpartition with axis=None'
        assert_equal(np.partition(a, kth, axis=None),
                     np.zeros_like(a.ravel(), dtype=np.intp), msg)

    def test_partition(self):
        d = np.arange(10)
        assert_raises(TypeError, np.partition, d, 2, kind=1)
        assert_raises(ValueError, np.partition, d, 2, kind="nonsense")
        assert_raises(ValueError, np.argpartition, d, 2, kind="nonsense")
        assert_raises(ValueError, d.partition, 2, axis=0, kind="nonsense")
        assert_raises(ValueError, d.argpartition, 2, axis=0, kind="nonsense")
        for k in ("introselect",):
            d = np.array([])
            assert_array_equal(np.partition(d, 0, kind=k), d)
            assert_array_equal(np.argpartition(d, 0, kind=k), d)
            d = np.ones(1)
            assert_array_equal(np.partition(d, 0, kind=k)[0], d)
            assert_array_equal(d[np.argpartition(d, 0, kind=k)],
                               np.partition(d, 0, kind=k))

            # kth not modified
            kth = np.array([30, 15, 5])
            okth = kth.copy()
            np.partition(np.arange(40), kth)
            assert_array_equal(kth, okth)

            for r in ([2, 1], [1, 2], [1, 1]):
                d = np.array(r)
                tgt = np.sort(d)
                assert_array_equal(np.partition(d, 0, kind=k)[0], tgt[0])
                assert_array_equal(np.partition(d, 1, kind=k)[1], tgt[1])
                self.assert_partitioned(np.partition(d, 0, kind=k), [0])
                self.assert_partitioned(d[np.argpartition(d, 0, kind=k)], [0])
                self.assert_partitioned(np.partition(d, 1, kind=k), [1])
                self.assert_partitioned(d[np.argpartition(d, 1, kind=k)], [1])
                for i in range(d.size):
                    d[i:].partition(0, kind=k)
                assert_array_equal(d, tgt)

            for r in ([3, 2, 1], [1, 2, 3], [2, 1, 3], [2, 3, 1],
                      [1, 1, 1], [1, 2, 2], [2, 2, 1], [1, 2, 1]):
                d = np.array(r)
                tgt = np.sort(d)
                assert_array_equal(np.partition(d, 0, kind=k)[0], tgt[0])
                assert_array_equal(np.partition(d, 1, kind=k)[1], tgt[1])
                assert_array_equal(np.partition(d, 2, kind=k)[2], tgt[2])
                self.assert_partitioned(np.partition(d, 0, kind=k), [0])
                self.assert_partitioned(d[np.argpartition(d, 0, kind=k)], [0])
                self.assert_partitioned(np.partition(d, 1, kind=k), [1])
                self.assert_partitioned(d[np.argpartition(d, 1, kind=k)], [1])
                self.assert_partitioned(np.partition(d, 2, kind=k), [2])
                self.assert_partitioned(d[np.argpartition(d, 2, kind=k)], [2])
                for i in range(d.size):
                    d[i:].partition(0, kind=k)
                assert_array_equal(d, tgt)

            d = np.ones(50)
            assert_array_equal(np.partition(d, 0, kind=k), d)
            assert_array_equal(d[np.argpartition(d, 0, kind=k)],
                               np.partition(d, 0, kind=k))

            # sorted
            d = np.arange(49)
            assert_equal(np.partition(d, 5, kind=k)[5], 5)
            assert_equal(np.partition(d, 15, kind=k)[15], 15)
            self.assert_partitioned(np.partition(d, 5, kind=k), [5])
            self.assert_partitioned(d[np.argpartition(d, 5, kind=k)], [5])
            self.assert_partitioned(np.partition(d, 15, kind=k), [15])
            self.assert_partitioned(d[np.argpartition(d, 15, kind=k)], [15])

            # rsorted
            d = np.arange(47)[::-1]
            assert_equal(np.partition(d, 6, kind=k)[6], 6)
            assert_equal(np.partition(d, 16, kind=k)[16], 16)
            self.assert_partitioned(np.partition(d, 6, kind=k), [6])
            self.assert_partitioned(d[np.argpartition(d, 6, kind=k)], [6])
            self.assert_partitioned(np.partition(d, 16, kind=k), [16])
            self.assert_partitioned(d[np.argpartition(d, 16, kind=k)], [16])

            assert_array_equal(np.partition(d, -6, kind=k),
                               np.partition(d, 41, kind=k))
            assert_array_equal(np.partition(d, -16, kind=k),
                               np.partition(d, 31, kind=k))
            self.assert_partitioned(np.partition(d, 41, kind=k), [41])
            self.assert_partitioned(d[np.argpartition(d, -6, kind=k)], [41])

            # median of 3 killer, O(n^2) on pure median 3 pivot quickselect
            # exercises the median of median of 5 code used to keep O(n)
            d = np.arange(1000000)
            x = np.roll(d, d.size // 2)
            mid = x.size // 2 + 1
            assert_equal(np.partition(x, mid)[mid], mid)
            d = np.arange(1000001)
            x = np.roll(d, d.size // 2 + 1)
            mid = x.size // 2 + 1
            assert_equal(np.partition(x, mid)[mid], mid)

            # max
            d = np.ones(10)
            d[1] = 4
            assert_equal(np.partition(d, (2, -1))[-1], 4)
            assert_equal(np.partition(d, (2, -1))[2], 1)
            assert_equal(d[np.argpartition(d, (2, -1))][-1], 4)
            assert_equal(d[np.argpartition(d, (2, -1))][2], 1)
            d[1] = np.nan
            assert_(np.isnan(d[np.argpartition(d, (2, -1))][-1]))
            assert_(np.isnan(np.partition(d, (2, -1))[-1]))

            # equal elements
            d = np.arange(47) % 7
            tgt = np.sort(np.arange(47) % 7)
            np.random.shuffle(d)
            for i in range(d.size):
                assert_equal(np.partition(d, i, kind=k)[i], tgt[i])
            self.assert_partitioned(np.partition(d, 6, kind=k), [6])
            self.assert_partitioned(d[np.argpartition(d, 6, kind=k)], [6])
            self.assert_partitioned(np.partition(d, 16, kind=k), [16])
            self.assert_partitioned(d[np.argpartition(d, 16, kind=k)], [16])
            for i in range(d.size):
                d[i:].partition(0, kind=k)
            assert_array_equal(d, tgt)

            d = np.array([0, 1, 2, 3, 4, 5, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7,
                          7, 7, 7, 7, 7, 9])
            kth = [0, 3, 19, 20]
            assert_equal(np.partition(d, kth, kind=k)[kth], (0, 3, 7, 7))
            assert_equal(d[np.argpartition(d, kth, kind=k)][kth], (0, 3, 7, 7))

            d = np.array([2, 1])
            d.partition(0, kind=k)
            assert_raises(ValueError, d.partition, 2)
            assert_raises(AxisError, d.partition, 3, axis=1)
            assert_raises(ValueError, np.partition, d, 2)
            assert_raises(AxisError, np.partition, d, 2, axis=1)
            assert_raises(ValueError, d.argpartition, 2)
            assert_raises(AxisError, d.argpartition, 3, axis=1)
            assert_raises(ValueError, np.argpartition, d, 2)
            assert_raises(AxisError, np.argpartition, d, 2, axis=1)
            d = np.arange(10).reshape((2, 5))
            d.partition(1, axis=0, kind=k)
            d.partition(4, axis=1, kind=k)
            np.partition(d, 1, axis=0, kind=k)
            np.partition(d, 4, axis=1, kind=k)
            np.partition(d, 1, axis=None, kind=k)
            np.partition(d, 9, axis=None, kind=k)
            d.argpartition(1, axis=0, kind=k)
            d.argpartition(4, axis=1, kind=k)
            np.argpartition(d, 1, axis=0, kind=k)
            np.argpartition(d, 4, axis=1, kind=k)
            np.argpartition(d, 1, axis=None, kind=k)
            np.argpartition(d, 9, axis=None, kind=k)
            assert_raises(ValueError, d.partition, 2, axis=0)
            assert_raises(ValueError, d.partition, 11, axis=1)
            assert_raises(TypeError, d.partition, 2, axis=None)
            assert_raises(ValueError, np.partition, d, 9, axis=1)
            assert_raises(ValueError, np.partition, d, 11, axis=None)
            assert_raises(ValueError, d.argpartition, 2, axis=0)
            assert_raises(ValueError, d.argpartition, 11, axis=1)
            assert_raises(ValueError, np.argpartition, d, 9, axis=1)
            assert_raises(ValueError, np.argpartition, d, 11, axis=None)

            td = [(dt, s) for dt in [np.int32, np.float32, np.complex64]
                  for s in (9, 16)]
            for dt, s in td:
                aae = assert_array_equal
                at = assert_

                d = np.arange(s, dtype=dt)
                np.random.shuffle(d)
                d1 = np.tile(np.arange(s, dtype=dt), (4, 1))
                map(np.random.shuffle, d1)
                d0 = np.transpose(d1)
                for i in range(d.size):
                    p = np.partition(d, i, kind=k)
                    assert_equal(p[i], i)
                    # all before are smaller
                    assert_array_less(p[:i], p[i])
                    # all after are larger
                    assert_array_less(p[i], p[i + 1:])
                    self.assert_partitioned(p, [i])
                    self.assert_partitioned(
                            d[np.argpartition(d, i, kind=k)], [i])

                    p = np.partition(d1, i, axis=1, kind=k)
                    parg = d1[np.arange(d1.shape[0])[:, None],
                            np.argpartition(d1, i, axis=1, kind=k)]
                    aae(p[:, i], np.array([i] * d1.shape[0], dtype=dt))
                    # array_less does not seem to work right
                    at((p[:, :i].T <= p[:, i]).all(),
                       msg="%d: %r <= %r" % (i, p[:, i], p[:, :i].T))
                    at((p[:, i + 1:].T > p[:, i]).all(),
                       msg="%d: %r < %r" % (i, p[:, i], p[:, i + 1:].T))
                    for row in range(p.shape[0]):
                        self.assert_partitioned(p[row], [i])
                        self.assert_partitioned(parg[row], [i])

                    p = np.partition(d0, i, axis=0, kind=k)
                    parg = d0[np.argpartition(d0, i, axis=0, kind=k),
                            np.arange(d0.shape[1])[None, :]]
                    aae(p[i, :], np.array([i] * d1.shape[0], dtype=dt))
                    # array_less does not seem to work right
                    at((p[:i, :] <= p[i, :]).all(),
                       msg="%d: %r <= %r" % (i, p[i, :], p[:i, :]))
                    at((p[i + 1:, :] > p[i, :]).all(),
                       msg="%d: %r < %r" % (i, p[i, :], p[:, i + 1:]))
                    for col in range(p.shape[1]):
                        self.assert_partitioned(p[:, col], [i])
                        self.assert_partitioned(parg[:, col], [i])

                    # check inplace
                    dc = d.copy()
                    dc.partition(i, kind=k)
                    assert_equal(dc, np.partition(d, i, kind=k))
                    dc = d0.copy()
                    dc.partition(i, axis=0, kind=k)
                    assert_equal(dc, np.partition(d0, i, axis=0, kind=k))
                    dc = d1.copy()
                    dc.partition(i, axis=1, kind=k)
                    assert_equal(dc, np.partition(d1, i, axis=1, kind=k))

    def assert_partitioned(self, d, kth):
        prev = 0
        for k in np.sort(kth):
            assert_array_compare(operator.__le__, d[prev:k], d[k],
                    err_msg='kth %d' % k)
            assert_((d[k:] >= d[k]).all(),
                    msg="kth %d, %r not greater equal %r" % (k, d[k:], d[k]))
            prev = k + 1

    def test_partition_iterative(self):
        d = np.arange(17)
        kth = (0, 1, 2, 429, 231)
        assert_raises(ValueError, d.partition, kth)
        assert_raises(ValueError, d.argpartition, kth)
        d = np.arange(10).reshape((2, 5))
        assert_raises(ValueError, d.partition, kth, axis=0)
        assert_raises(ValueError, d.partition, kth, axis=1)
        assert_raises(ValueError, np.partition, d, kth, axis=1)
        assert_raises(ValueError, np.partition, d, kth, axis=None)

        d = np.array([3, 4, 2, 1])
        p = np.partition(d, (0, 3))
        self.assert_partitioned(p, (0, 3))
        self.assert_partitioned(d[np.argpartition(d, (0, 3))], (0, 3))

        assert_array_equal(p, np.partition(d, (-3, -1)))
        assert_array_equal(p, d[np.argpartition(d, (-3, -1))])

        d = np.arange(17)
        np.random.shuffle(d)
        d.partition(range(d.size))
        assert_array_equal(np.arange(17), d)
        np.random.shuffle(d)
        assert_array_equal(np.arange(17), d[d.argpartition(range(d.size))])

        # test unsorted kth
        d = np.arange(17)
        np.random.shuffle(d)
        keys = np.array([1, 3, 8, -2])
        np.random.shuffle(d)
        p = np.partition(d, keys)
        self.assert_partitioned(p, keys)
        p = d[np.argpartition(d, keys)]
        self.assert_partitioned(p, keys)
        np.random.shuffle(keys)
        assert_array_equal(np.partition(d, keys), p)
        assert_array_equal(d[np.argpartition(d, keys)], p)

        # equal kth
        d = np.arange(20)[::-1]
        self.assert_partitioned(np.partition(d, [5] * 4), [5])
        self.assert_partitioned(np.partition(d, [5] * 4 + [6, 13]),
                                [5] * 4 + [6, 13])
        self.assert_partitioned(d[np.argpartition(d, [5] * 4)], [5])
        self.assert_partitioned(d[np.argpartition(d, [5] * 4 + [6, 13])],
                                [5] * 4 + [6, 13])

        d = np.arange(12)
        np.random.shuffle(d)
        d1 = np.tile(np.arange(12), (4, 1))
        map(np.random.shuffle, d1)
        d0 = np.transpose(d1)

        kth = (1, 6, 7, -1)
        p = np.partition(d1, kth, axis=1)
        pa = d1[np.arange(d1.shape[0])[:, None],
                d1.argpartition(kth, axis=1)]
        assert_array_equal(p, pa)
        for i in range(d1.shape[0]):
            self.assert_partitioned(p[i, :], kth)
        p = np.partition(d0, kth, axis=0)
        pa = d0[np.argpartition(d0, kth, axis=0),
                np.arange(d0.shape[1])[None, :]]
        assert_array_equal(p, pa)
        for i in range(d0.shape[1]):
            self.assert_partitioned(p[:, i], kth)

    def test_partition_cdtype(self):
        d = np.array([('Galahad', 1.7, 38), ('Arthur', 1.8, 41),
                   ('Lancelot', 1.9, 38)],
                  dtype=[('name', '|S10'), ('height', '<f8'), ('age', '<i4')])

        tgt = np.sort(d, order=['age', 'height'])
        assert_array_equal(np.partition(d, range(d.size),
                                        order=['age', 'height']),
                           tgt)
        assert_array_equal(d[np.argpartition(d, range(d.size),
                                             order=['age', 'height'])],
                           tgt)
        for k in range(d.size):
            assert_equal(np.partition(d, k, order=['age', 'height'])[k],
                        tgt[k])
            assert_equal(d[np.argpartition(d, k, order=['age', 'height'])][k],
                         tgt[k])

        d = np.array(['Galahad', 'Arthur', 'zebra', 'Lancelot'])
        tgt = np.sort(d)
        assert_array_equal(np.partition(d, range(d.size)), tgt)
        for k in range(d.size):
            assert_equal(np.partition(d, k)[k], tgt[k])
            assert_equal(d[np.argpartition(d, k)][k], tgt[k])

    def test_partition_unicode_kind(self):
        d = np.arange(10)
        k = b'\xc3\xa4'.decode("UTF8")
        assert_raises(ValueError, d.partition, 2, kind=k)
        assert_raises(ValueError, d.argpartition, 2, kind=k)

    def test_partition_fuzz(self):
        # a few rounds of random data testing
        for j in range(10, 30):
            for i in range(1, j - 2):
                d = np.arange(j)
                np.random.shuffle(d)
                d = d % np.random.randint(2, 30)
                idx = np.random.randint(d.size)
                kth = [0, idx, i, i + 1]
                tgt = np.sort(d)[kth]
                assert_array_equal(np.partition(d, kth)[kth], tgt,
                                   err_msg=f"data: {d!r}\n kth: {kth!r}")

    @pytest.mark.parametrize("kth_dtype", np.typecodes["AllInteger"])
    def test_argpartition_gh5524(self, kth_dtype):
        #  A test for functionality of argpartition on lists.
        kth = np.array(1, dtype=kth_dtype)[()]
        d = [6, 7, 3, 2, 9, 0]
        p = np.argpartition(d, kth)
        self.assert_partitioned(np.array(d)[p], [1])

    def test_flatten(self):
        x0 = np.array([[1, 2, 3], [4, 5, 6]], np.int32)
        x1 = np.array([[[1, 2], [3, 4]], [[5, 6], [7, 8]]], np.int32)
        y0 = np.array([1, 2, 3, 4, 5, 6], np.int32)
        y0f = np.array([1, 4, 2, 5, 3, 6], np.int32)
        y1 = np.array([1, 2, 3, 4, 5, 6, 7, 8], np.int32)
        y1f = np.array([1, 5, 3, 7, 2, 6, 4, 8], np.int32)
        assert_equal(x0.flatten(), y0)
        assert_equal(x0.flatten('F'), y0f)
        assert_equal(x0.flatten('F'), x0.T.flatten())
        assert_equal(x1.flatten(), y1)
        assert_equal(x1.flatten('F'), y1f)
        assert_equal(x1.flatten('F'), x1.T.flatten())

    @pytest.mark.parametrize('func', (np.dot, np.matmul))
    def test_arr_mult(self, func):
        a = np.array([[1, 0], [0, 1]])
        b = np.array([[0, 1], [1, 0]])
        c = np.array([[9, 1], [1, -9]])
        d = np.arange(24).reshape(4, 6)
        ddt = np.array(
            [[  55,  145,  235,  325],
             [ 145,  451,  757, 1063],
             [ 235,  757, 1279, 1801],
             [ 325, 1063, 1801, 2539]]
        )
        dtd = np.array(
            [[504, 540, 576, 612, 648, 684],
             [540, 580, 620, 660, 700, 740],
             [576, 620, 664, 708, 752, 796],
             [612, 660, 708, 756, 804, 852],
             [648, 700, 752, 804, 856, 908],
             [684, 740, 796, 852, 908, 964]]
        )

        # gemm vs syrk optimizations
        for et in [np.float32, np.float64, np.complex64, np.complex128]:
            eaf = a.astype(et)
            assert_equal(func(eaf, eaf), eaf)
            assert_equal(func(eaf.T, eaf), eaf)
            assert_equal(func(eaf, eaf.T), eaf)
            assert_equal(func(eaf.T, eaf.T), eaf)
            assert_equal(func(eaf.T.copy(), eaf), eaf)
            assert_equal(func(eaf, eaf.T.copy()), eaf)
            assert_equal(func(eaf.T.copy(), eaf.T.copy()), eaf)

        # syrk validations
        for et in [np.float32, np.float64, np.complex64, np.complex128]:
            eaf = a.astype(et)
            ebf = b.astype(et)
            assert_equal(func(ebf, ebf), eaf)
            assert_equal(func(ebf.T, ebf), eaf)
            assert_equal(func(ebf, ebf.T), eaf)
            assert_equal(func(ebf.T, ebf.T), eaf)

        # syrk - different shape, stride, and view validations
        for et in [np.float32, np.float64, np.complex64, np.complex128]:
            edf = d.astype(et)
            assert_equal(
                func(edf[::-1, :], edf.T),
                func(edf[::-1, :].copy(), edf.T.copy())
            )
            assert_equal(
                func(edf[:, ::-1], edf.T),
                func(edf[:, ::-1].copy(), edf.T.copy())
            )
            assert_equal(
                func(edf, edf[::-1, :].T),
                func(edf, edf[::-1, :].T.copy())
            )
            assert_equal(
                func(edf, edf[:, ::-1].T),
                func(edf, edf[:, ::-1].T.copy())
            )
            assert_equal(
                func(edf[:edf.shape[0] // 2, :], edf[::2, :].T),
                func(edf[:edf.shape[0] // 2, :].copy(), edf[::2, :].T.copy())
            )
            assert_equal(
                func(edf[::2, :], edf[:edf.shape[0] // 2, :].T),
                func(edf[::2, :].copy(), edf[:edf.shape[0] // 2, :].T.copy())
            )

        # syrk - different shape
        for et in [np.float32, np.float64, np.complex64, np.complex128]:
            edf = d.astype(et)
            eddtf = ddt.astype(et)
            edtdf = dtd.astype(et)
            assert_equal(func(edf, edf.T), eddtf)
            assert_equal(func(edf.T, edf), edtdf)

    @pytest.mark.parametrize('func', (np.dot, np.matmul))
    @pytest.mark.parametrize('dtype', 'ifdFD')
    def test_no_dgemv(self, func, dtype):
        # check vector arg for contiguous before gemv
        # gh-12156
        a = np.arange(8.0, dtype=dtype).reshape(2, 4)
        b = np.broadcast_to(1., (4, 1))
        ret1 = func(a, b)
        ret2 = func(a, b.copy())
        assert_equal(ret1, ret2)

        ret1 = func(b.T, a.T)
        ret2 = func(b.T.copy(), a.T)
        assert_equal(ret1, ret2)

        # check for unaligned data
        dt = np.dtype(dtype)
        a = np.zeros(8 * dt.itemsize // 2 + 1, dtype='int16')[1:].view(dtype)
        a = a.reshape(2, 4)
        b = a[0]
        # make sure it is not aligned
        assert_(a.__array_interface__['data'][0] % dt.itemsize != 0)
        ret1 = func(a, b)
        ret2 = func(a.copy(), b.copy())
        assert_equal(ret1, ret2)

        ret1 = func(b.T, a.T)
        ret2 = func(b.T.copy(), a.T.copy())
        assert_equal(ret1, ret2)

    def test_dot(self):
        a = np.array([[1, 0], [0, 1]])
        b = np.array([[0, 1], [1, 0]])
        c = np.array([[9, 1], [1, -9]])
        # function versus methods
        assert_equal(np.dot(a, b), a.dot(b))
        assert_equal(np.dot(np.dot(a, b), c), a.dot(b).dot(c))

        # test passing in an output array
        c = np.zeros_like(a)
        a.dot(b, c)
        assert_equal(c, np.dot(a, b))

        # test keyword args
        c = np.zeros_like(a)
        a.dot(b=b, out=c)
        assert_equal(c, np.dot(a, b))

    @pytest.mark.parametrize("dtype", [np.half, np.double, np.longdouble])
    @pytest.mark.skipif(IS_WASM, reason="no wasm fp exception support")
    def test_dot_errstate(self, dtype):
        a = np.array([1, 1], dtype=dtype)
        b = np.array([-np.inf, np.inf], dtype=dtype)

        with np.errstate(invalid='raise'):
            # there are two paths, depending on the number of dimensions - test
            # them both
            with pytest.raises(FloatingPointError,
                    match="invalid value encountered in dot"):
                np.dot(a, b)

            # test that fp exceptions are properly cleared
            np.dot(a, a)

            with pytest.raises(FloatingPointError,
                    match="invalid value encountered in dot"):
                np.dot(a[np.newaxis, np.newaxis, ...],
                       b[np.newaxis, ..., np.newaxis])

            np.dot(a[np.newaxis, np.newaxis, ...],
                   a[np.newaxis, ..., np.newaxis])

    def test_dot_type_mismatch(self):
        c = 1.
        A = np.array((1, 1), dtype='i,i')

        assert_raises(TypeError, np.dot, c, A)
        assert_raises(TypeError, np.dot, A, c)

    def test_dot_out_mem_overlap(self):
        np.random.seed(1)

        # Test BLAS and non-BLAS code paths, including all dtypes
        # that dot() supports
        dtypes = [np.dtype(code) for code in np.typecodes['All']
                  if code not in 'USVM']
        for dtype in dtypes:
            a = np.random.rand(3, 3).astype(dtype)

            # Valid dot() output arrays must be aligned
            b = _aligned_zeros((3, 3), dtype=dtype)
            b[...] = np.random.rand(3, 3)

            y = np.dot(a, b)
            x = np.dot(a, b, out=b)
            assert_equal(x, y, err_msg=repr(dtype))

            # Check invalid output array
            assert_raises(ValueError, np.dot, a, b, out=b[::2])
            assert_raises(ValueError, np.dot, a, b, out=b.T)

    def test_dot_matmul_out(self):
        # gh-9641
        class Sub(np.ndarray):
            pass
        a = np.ones((2, 2)).view(Sub)
        b = np.ones((2, 2)).view(Sub)
        out = np.ones((2, 2))

        # make sure out can be any ndarray (not only subclass of inputs)
        np.dot(a, b, out=out)
        np.matmul(a, b, out=out)

    def test_dot_matmul_inner_array_casting_fails(self):

        class A:
            def __array__(self, *args, **kwargs):
                raise NotImplementedError

        # Don't override the error from calling __array__()
        assert_raises(NotImplementedError, np.dot, A(), A())
        assert_raises(NotImplementedError, np.matmul, A(), A())
        assert_raises(NotImplementedError, np.inner, A(), A())

    def test_matmul_out(self):
        # overlapping memory
        a = np.arange(18).reshape(2, 3, 3)
        b = np.matmul(a, a)
        c = np.matmul(a, a, out=a)
        assert_(c is a)
        assert_equal(c, b)
        a = np.arange(18).reshape(2, 3, 3)
        c = np.matmul(a, a, out=a[::-1, ...])
        assert_(c.base is a.base)
        assert_equal(c, b)

    def test_diagonal(self):
        a = np.arange(12).reshape((3, 4))
        assert_equal(a.diagonal(), [0, 5, 10])
        assert_equal(a.diagonal(0), [0, 5, 10])
        assert_equal(a.diagonal(1), [1, 6, 11])
        assert_equal(a.diagonal(-1), [4, 9])
        assert_raises(AxisError, a.diagonal, axis1=0, axis2=5)
        assert_raises(AxisError, a.diagonal, axis1=5, axis2=0)
        assert_raises(AxisError, a.diagonal, axis1=5, axis2=5)
        assert_raises(ValueError, a.diagonal, axis1=1, axis2=1)

        b = np.arange(8).reshape((2, 2, 2))
        assert_equal(b.diagonal(), [[0, 6], [1, 7]])
        assert_equal(b.diagonal(0), [[0, 6], [1, 7]])
        assert_equal(b.diagonal(1), [[2], [3]])
        assert_equal(b.diagonal(-1), [[4], [5]])
        assert_raises(ValueError, b.diagonal, axis1=0, axis2=0)
        assert_equal(b.diagonal(0, 1, 2), [[0, 3], [4, 7]])
        assert_equal(b.diagonal(0, 0, 1), [[0, 6], [1, 7]])
        assert_equal(b.diagonal(offset=1, axis1=0, axis2=2), [[1], [3]])
        # Order of axis argument doesn't matter:
        assert_equal(b.diagonal(0, 2, 1), [[0, 3], [4, 7]])

    def test_diagonal_view_notwriteable(self):
        a = np.eye(3).diagonal()
        assert_(not a.flags.writeable)
        assert_(not a.flags.owndata)

        a = np.diagonal(np.eye(3))
        assert_(not a.flags.writeable)
        assert_(not a.flags.owndata)

        a = np.diag(np.eye(3))
        assert_(not a.flags.writeable)
        assert_(not a.flags.owndata)

    def test_diagonal_memleak(self):
        # Regression test for a bug that crept in at one point
        a = np.zeros((100, 100))
        if HAS_REFCOUNT:
            assert_(sys.getrefcount(a) < 50)
        for i in range(100):
            a.diagonal()
        if HAS_REFCOUNT:
            assert_(sys.getrefcount(a) < 50)

    def test_size_zero_memleak(self):
        # Regression test for issue 9615
        # Exercises a special-case code path for dot products of length
        # zero in cblasfuncs (making it is specific to floating dtypes).
        a = np.array([], dtype=np.float64)
        x = np.array(2.0)
        for _ in range(100):
            np.dot(a, a, out=x)
        if HAS_REFCOUNT:
            assert_(sys.getrefcount(x) < 50)

    def test_trace(self):
        a = np.arange(12).reshape((3, 4))
        assert_equal(a.trace(), 15)
        assert_equal(a.trace(0), 15)
        assert_equal(a.trace(1), 18)
        assert_equal(a.trace(-1), 13)

        b = np.arange(8).reshape((2, 2, 2))
        assert_equal(b.trace(), [6, 8])
        assert_equal(b.trace(0), [6, 8])
        assert_equal(b.trace(1), [2, 3])
        assert_equal(b.trace(-1), [4, 5])
        assert_equal(b.trace(0, 0, 1), [6, 8])
        assert_equal(b.trace(0, 0, 2), [5, 9])
        assert_equal(b.trace(0, 1, 2), [3, 11])
        assert_equal(b.trace(offset=1, axis1=0, axis2=2), [1, 3])

        out = np.array(1)
        ret = a.trace(out=out)
        assert ret is out

    def test_trace_subclass(self):
        # The class would need to overwrite trace to ensure single-element
        # output also has the right subclass.
        class MyArray(np.ndarray):
            pass

        b = np.arange(8).reshape((2, 2, 2)).view(MyArray)
        t = b.trace()
        assert_(isinstance(t, MyArray))

    def test_put(self):
        icodes = np.typecodes['AllInteger']
        fcodes = np.typecodes['AllFloat']
        for dt in icodes + fcodes + 'O':
            tgt = np.array([0, 1, 0, 3, 0, 5], dtype=dt)

            # test 1-d
            a = np.zeros(6, dtype=dt)
            a.put([1, 3, 5], [1, 3, 5])
            assert_equal(a, tgt)

            # test 2-d
            a = np.zeros((2, 3), dtype=dt)
            a.put([1, 3, 5], [1, 3, 5])
            assert_equal(a, tgt.reshape(2, 3))

        for dt in '?':
            tgt = np.array([False, True, False, True, False, True], dtype=dt)

            # test 1-d
            a = np.zeros(6, dtype=dt)
            a.put([1, 3, 5], [True] * 3)
            assert_equal(a, tgt)

            # test 2-d
            a = np.zeros((2, 3), dtype=dt)
            a.put([1, 3, 5], [True] * 3)
            assert_equal(a, tgt.reshape(2, 3))

        # check must be writeable
        a = np.zeros(6)
        a.flags.writeable = False
        assert_raises(ValueError, a.put, [1, 3, 5], [1, 3, 5])

        # when calling np.put, make sure a
        # TypeError is raised if the object
        # isn't an ndarray
        bad_array = [1, 2, 3]
        assert_raises(TypeError, np.put, bad_array, [0, 2], 5)

        # when calling np.put, make sure an
        # IndexError is raised if the
        # array is empty
        empty_array = np.asarray([])
        with pytest.raises(IndexError,
                            match="cannot replace elements of an empty array"):
            np.put(empty_array, 1, 1, mode="wrap")
        with pytest.raises(IndexError,
                            match="cannot replace elements of an empty array"):
            np.put(empty_array, 1, 1, mode="clip")

    def test_ravel(self):
        a = np.array([[0, 1], [2, 3]])
        assert_equal(a.ravel(), [0, 1, 2, 3])
        assert_(not a.ravel().flags.owndata)
        assert_equal(a.ravel('F'), [0, 2, 1, 3])
        assert_equal(a.ravel(order='C'), [0, 1, 2, 3])
        assert_equal(a.ravel(order='F'), [0, 2, 1, 3])
        assert_equal(a.ravel(order='A'), [0, 1, 2, 3])
        assert_(not a.ravel(order='A').flags.owndata)
        assert_equal(a.ravel(order='K'), [0, 1, 2, 3])
        assert_(not a.ravel(order='K').flags.owndata)
        assert_equal(a.ravel(), a.reshape(-1))

        a = np.array([[0, 1], [2, 3]], order='F')
        assert_equal(a.ravel(), [0, 1, 2, 3])
        assert_equal(a.ravel(order='A'), [0, 2, 1, 3])
        assert_equal(a.ravel(order='K'), [0, 2, 1, 3])
        assert_(not a.ravel(order='A').flags.owndata)
        assert_(not a.ravel(order='K').flags.owndata)
        assert_equal(a.ravel(), a.reshape(-1))
        assert_equal(a.ravel(order='A'), a.reshape(-1, order='A'))

        a = np.array([[0, 1], [2, 3]])[::-1, :]
        assert_equal(a.ravel(), [2, 3, 0, 1])
        assert_equal(a.ravel(order='C'), [2, 3, 0, 1])
        assert_equal(a.ravel(order='F'), [2, 0, 3, 1])
        assert_equal(a.ravel(order='A'), [2, 3, 0, 1])
        # 'K' doesn't reverse the axes of negative strides
        assert_equal(a.ravel(order='K'), [2, 3, 0, 1])
        assert_(a.ravel(order='K').flags.owndata)

        # Test simple 1-d copy behaviour:
        a = np.arange(10)[::2]
        assert_(a.ravel('K').flags.owndata)
        assert_(a.ravel('C').flags.owndata)
        assert_(a.ravel('F').flags.owndata)

        # Not contiguous and 1-sized axis with non matching stride
        a = np.arange(2**3 * 2)[::2]
        a = a.reshape(2, 1, 2, 2).swapaxes(-1, -2)
        strides = list(a.strides)
        strides[1] = 123
        a.strides = strides
        assert_(a.ravel(order='K').flags.owndata)
        assert_equal(a.ravel('K'), np.arange(0, 15, 2))

        # contiguous and 1-sized axis with non matching stride works:
        a = np.arange(2**3)
        a = a.reshape(2, 1, 2, 2).swapaxes(-1, -2)
        strides = list(a.strides)
        strides[1] = 123
        a.strides = strides
        assert_(np.may_share_memory(a.ravel(order='K'), a))
        assert_equal(a.ravel(order='K'), np.arange(2**3))

        # Test negative strides (not very interesting since non-contiguous):
        a = np.arange(4)[::-1].reshape(2, 2)
        assert_(a.ravel(order='C').flags.owndata)
        assert_(a.ravel(order='K').flags.owndata)
        assert_equal(a.ravel('C'), [3, 2, 1, 0])
        assert_equal(a.ravel('K'), [3, 2, 1, 0])

        # 1-element tidy strides test:
        a = np.array([[1]])
        a.strides = (123, 432)
        if np.ones(1).strides == (8,):
            assert_(np.may_share_memory(a.ravel('K'), a))
            assert_equal(a.ravel('K').strides, (a.dtype.itemsize,))

        for order in ('C', 'F', 'A', 'K'):
            # 0-d corner case:
            a = np.array(0)
            assert_equal(a.ravel(order), [0])
            assert_(np.may_share_memory(a.ravel(order), a))

        # Test that certain non-inplace ravels work right (mostly) for 'K':
        b = np.arange(2**4 * 2)[::2].reshape(2, 2, 2, 2)
        a = b[..., ::2]
        assert_equal(a.ravel('K'), [0, 4, 8, 12, 16, 20, 24, 28])
        assert_equal(a.ravel('C'), [0, 4, 8, 12, 16, 20, 24, 28])
        assert_equal(a.ravel('A'), [0, 4, 8, 12, 16, 20, 24, 28])
        assert_equal(a.ravel('F'), [0, 16, 8, 24, 4, 20, 12, 28])

        a = b[::2, ...]
        assert_equal(a.ravel('K'), [0, 2, 4, 6, 8, 10, 12, 14])
        assert_equal(a.ravel('C'), [0, 2, 4, 6, 8, 10, 12, 14])
        assert_equal(a.ravel('A'), [0, 2, 4, 6, 8, 10, 12, 14])
        assert_equal(a.ravel('F'), [0, 8, 4, 12, 2, 10, 6, 14])

    def test_ravel_subclass(self):
        class ArraySubclass(np.ndarray):
            pass

        a = np.arange(10).view(ArraySubclass)
        assert_(isinstance(a.ravel('C'), ArraySubclass))
        assert_(isinstance(a.ravel('F'), ArraySubclass))
        assert_(isinstance(a.ravel('A'), ArraySubclass))
        assert_(isinstance(a.ravel('K'), ArraySubclass))

        a = np.arange(10)[::2].view(ArraySubclass)
        assert_(isinstance(a.ravel('C'), ArraySubclass))
        assert_(isinstance(a.ravel('F'), ArraySubclass))
        assert_(isinstance(a.ravel('A'), ArraySubclass))
        assert_(isinstance(a.ravel('K'), ArraySubclass))

    def test_swapaxes(self):
        a = np.arange(1 * 2 * 3 * 4).reshape(1, 2, 3, 4).copy()
        idx = np.indices(a.shape)
        assert_(a.flags['OWNDATA'])
        b = a.copy()
        # check exceptions
        assert_raises(AxisError, a.swapaxes, -5, 0)
        assert_raises(AxisError, a.swapaxes, 4, 0)
        assert_raises(AxisError, a.swapaxes, 0, -5)
        assert_raises(AxisError, a.swapaxes, 0, 4)

        for i in range(-4, 4):
            for j in range(-4, 4):
                for k, src in enumerate((a, b)):
                    c = src.swapaxes(i, j)
                    # check shape
                    shape = list(src.shape)
                    shape[i] = src.shape[j]
                    shape[j] = src.shape[i]
                    assert_equal(c.shape, shape, str((i, j, k)))
                    # check array contents
                    i0, i1, i2, i3 = [dim - 1 for dim in c.shape]
                    j0, j1, j2, j3 = [dim - 1 for dim in src.shape]
                    assert_equal(src[idx[j0], idx[j1], idx[j2], idx[j3]],
                                 c[idx[i0], idx[i1], idx[i2], idx[i3]],
                                 str((i, j, k)))
                    # check a view is always returned, gh-5260
                    assert_(not c.flags['OWNDATA'], str((i, j, k)))
                    # check on non-contiguous input array
                    if k == 1:
                        b = c

    def test_conjugate(self):
        a = np.array([1 - 1j, 1 + 1j, 23 + 23.0j])
        ac = a.conj()
        assert_equal(a.real, ac.real)
        assert_equal(a.imag, -ac.imag)
        assert_equal(ac, a.conjugate())
        assert_equal(ac, np.conjugate(a))

        a = np.array([1 - 1j, 1 + 1j, 23 + 23.0j], 'F')
        ac = a.conj()
        assert_equal(a.real, ac.real)
        assert_equal(a.imag, -ac.imag)
        assert_equal(ac, a.conjugate())
        assert_equal(ac, np.conjugate(a))

        a = np.array([1, 2, 3])
        ac = a.conj()
        assert_equal(a, ac)
        assert_equal(ac, a.conjugate())
        assert_equal(ac, np.conjugate(a))

        a = np.array([1.0, 2.0, 3.0])
        ac = a.conj()
        assert_equal(a, ac)
        assert_equal(ac, a.conjugate())
        assert_equal(ac, np.conjugate(a))

        a = np.array([1 - 1j, 1 + 1j, 1, 2.0], object)
        ac = a.conj()
        assert_equal(ac, [k.conjugate() for k in a])
        assert_equal(ac, a.conjugate())
        assert_equal(ac, np.conjugate(a))

        a = np.array([1 - 1j, 1, 2.0, 'f'], object)
        assert_raises(TypeError, a.conj)
        assert_raises(TypeError, a.conjugate)

    def test_conjugate_out(self):
        # Minimal test for the out argument being passed on correctly
        # NOTE: The ability to pass `out` is currently undocumented!
        a = np.array([1 - 1j, 1 + 1j, 23 + 23.0j])
        out = np.empty_like(a)
        res = a.conjugate(out)
        assert res is out
        assert_array_equal(out, a.conjugate())

    def test_conjugate_scalar(self):
        for v in 5, 5j:
            a = np.array(v)
            assert a.conjugate() == v.conjugate()
        for a in (np.array('s'), np.array('2016', 'M'),
                np.array((1, 2), [('a', int), ('b', int)])):
            with pytest.raises(TypeError):
                a.conjugate()

    def test__complex__(self):
        dtypes = ['i1', 'i2', 'i4', 'i8',
                  'u1', 'u2', 'u4', 'u8',
                  'f', 'd', 'g', 'F', 'D', 'G',
                  '?', 'O']
        for dt in dtypes:
            a = np.array(7, dtype=dt)
            b = np.array([7], dtype=dt)
            c = np.array([[[[[7]]]]], dtype=dt)

            msg = f'dtype: {dt}'
            ap = complex(a)
            assert_equal(ap, a, msg)

            with assert_warns(DeprecationWarning):
                bp = complex(b)
            assert_equal(bp, b, msg)

            with assert_warns(DeprecationWarning):
                cp = complex(c)
            assert_equal(cp, c, msg)

    def test__complex__should_not_work(self):
        dtypes = ['i1', 'i2', 'i4', 'i8',
                  'u1', 'u2', 'u4', 'u8',
                  'f', 'd', 'g', 'F', 'D', 'G',
                  '?', 'O']
        for dt in dtypes:
            a = np.array([1, 2, 3], dtype=dt)
            assert_raises(TypeError, complex, a)

        dt = np.dtype([('a', 'f8'), ('b', 'i1')])
        b = np.array((1.0, 3), dtype=dt)
        assert_raises(TypeError, complex, b)

        c = np.array([(1.0, 3), (2e-3, 7)], dtype=dt)
        assert_raises(TypeError, complex, c)

        d = np.array('1+1j')
        assert_raises(TypeError, complex, d)

        e = np.array(['1+1j'], 'U')
        with assert_warns(DeprecationWarning):
            assert_raises(TypeError, complex, e)

class TestCequenceMethods:
    def test_array_contains(self):
        assert_(4.0 in np.arange(16.).reshape(4, 4))
        assert_(20.0 not in np.arange(16.).reshape(4, 4))

class TestBinop:
    def test_inplace(self):
        # test refcount 1 inplace conversion
        assert_array_almost_equal(np.array([0.5]) * np.array([1.0, 2.0]),
                                  [0.5, 1.0])

        d = np.array([0.5, 0.5])[::2]
        assert_array_almost_equal(d * (d * np.array([1.0, 2.0])),
                                  [0.25, 0.5])

        a = np.array([0.5])
        b = np.array([0.5])
        c = a + b
        c = a - b
        c = a * b
        c = a / b
        assert_equal(a, b)
        assert_almost_equal(c, 1.)

        c = a + b * 2. / b * a - a / b
        assert_equal(a, b)
        assert_equal(c, 0.5)

        # true divide
        a = np.array([5])
        b = np.array([3])
        c = (a * a) / b

        assert_almost_equal(c, 25 / 3)
        assert_equal(a, 5)
        assert_equal(b, 3)

    # ndarray.__rop__ always calls ufunc
    # ndarray.__iop__ always calls ufunc
    # ndarray.__op__, __rop__:
    #   - defer if other has __array_ufunc__ and it is None
    #           or other is not a subclass and has higher array priority
    #   - else, call ufunc
    @pytest.mark.xfail(IS_PYPY, reason="Bug in pypy3.{9, 10}-v7.3.13, #24862")
    def test_ufunc_binop_interaction(self):
        # Python method name (without underscores)
        #   -> (numpy ufunc, has_in_place_version, preferred_dtype)
        ops = {
            'add':      (np.add, True, float),
            'sub':      (np.subtract, True, float),
            'mul':      (np.multiply, True, float),
            'truediv':  (np.true_divide, True, float),
            'floordiv': (np.floor_divide, True, float),
            'mod':      (np.remainder, True, float),
            'divmod':   (np.divmod, False, float),
            'pow':      (np.power, True, int),
            'lshift':   (np.left_shift, True, int),
            'rshift':   (np.right_shift, True, int),
            'and':      (np.bitwise_and, True, int),
            'xor':      (np.bitwise_xor, True, int),
            'or':       (np.bitwise_or, True, int),
            'matmul':   (np.matmul, True, float),
            # 'ge':       (np.less_equal, False),
            # 'gt':       (np.less, False),
            # 'le':       (np.greater_equal, False),
            # 'lt':       (np.greater, False),
            # 'eq':       (np.equal, False),
            # 'ne':       (np.not_equal, False),
        }

        class Coerced(Exception):
            pass

        def array_impl(self):
            raise Coerced

        def op_impl(self, other):
            return "forward"

        def rop_impl(self, other):
            return "reverse"

        def iop_impl(self, other):
            return "in-place"

        def array_ufunc_impl(self, ufunc, method, *args, **kwargs):
            return ("__array_ufunc__", ufunc, method, args, kwargs)

        # Create an object with the given base, in the given module, with a
        # bunch of placeholder __op__ methods, and optionally a
        # __array_ufunc__ and __array_priority__.
        def make_obj(base, array_priority=False, array_ufunc=False,
                     alleged_module="__main__"):
            class_namespace = {"__array__": array_impl}
            if array_priority is not False:
                class_namespace["__array_priority__"] = array_priority
            for op in ops:
                class_namespace[f"__{op}__"] = op_impl
                class_namespace[f"__r{op}__"] = rop_impl
                class_namespace[f"__i{op}__"] = iop_impl
            if array_ufunc is not False:
                class_namespace["__array_ufunc__"] = array_ufunc
            eval_namespace = {"base": base,
                              "class_namespace": class_namespace,
                              "__name__": alleged_module,
                              }
            MyType = eval("type('MyType', (base,), class_namespace)",
                          eval_namespace)
            if issubclass(MyType, np.ndarray):
                # Use this range to avoid special case weirdnesses around
                # divide-by-0, pow(x, 2), overflow due to pow(big, big), etc.
                return np.arange(3, 7).reshape(2, 2).view(MyType)
            else:
                return MyType()

        def check(obj, binop_override_expected, ufunc_override_expected,
                  inplace_override_expected, check_scalar=True):
            for op, (ufunc, has_inplace, dtype) in ops.items():
                err_msg = ('op: %s, ufunc: %s, has_inplace: %s, dtype: %s'
                           % (op, ufunc, has_inplace, dtype))
                check_objs = [np.arange(3, 7, dtype=dtype).reshape(2, 2)]
                if check_scalar:
                    check_objs.append(check_objs[0][0])
                for arr in check_objs:
                    arr_method = getattr(arr, f"__{op}__")

                    def first_out_arg(result):
                        if op == "divmod":
                            assert_(isinstance(result, tuple))
                            return result[0]
                        else:
                            return result

                    # arr __op__ obj
                    if binop_override_expected:
                        assert_equal(arr_method(obj), NotImplemented, err_msg)
                    elif ufunc_override_expected:
                        assert_equal(arr_method(obj)[0], "__array_ufunc__",
                                     err_msg)
                    elif (isinstance(obj, np.ndarray) and
                        (type(obj).__array_ufunc__ is
                         np.ndarray.__array_ufunc__)):
                        # __array__ gets ignored
                        res = first_out_arg(arr_method(obj))
                        assert_(res.__class__ is obj.__class__, err_msg)
                    else:
                        assert_raises((TypeError, Coerced),
                                      arr_method, obj, err_msg=err_msg)
                    # obj __op__ arr
                    arr_rmethod = getattr(arr, f"__r{op}__")
                    if ufunc_override_expected:
                        res = arr_rmethod(obj)
                        assert_equal(res[0], "__array_ufunc__",
                                     err_msg=err_msg)
                        assert_equal(res[1], ufunc, err_msg=err_msg)
                    elif (isinstance(obj, np.ndarray) and
                            (type(obj).__array_ufunc__ is
                             np.ndarray.__array_ufunc__)):
                        # __array__ gets ignored
                        res = first_out_arg(arr_rmethod(obj))
                        assert_(res.__class__ is obj.__class__, err_msg)
                    else:
                        # __array_ufunc__ = "asdf" creates a TypeError
                        assert_raises((TypeError, Coerced),
                                      arr_rmethod, obj, err_msg=err_msg)

                    # arr __iop__ obj
                    # array scalars don't have in-place operators
                    if has_inplace and isinstance(arr, np.ndarray):
                        arr_imethod = getattr(arr, f"__i{op}__")
                        if inplace_override_expected:
                            assert_equal(arr_method(obj), NotImplemented,
                                         err_msg=err_msg)
                        elif ufunc_override_expected:
                            res = arr_imethod(obj)
                            assert_equal(res[0], "__array_ufunc__", err_msg)
                            assert_equal(res[1], ufunc, err_msg)
                            assert_(type(res[-1]["out"]) is tuple, err_msg)
                            assert_(res[-1]["out"][0] is arr, err_msg)
                        elif (isinstance(obj, np.ndarray) and
                                (type(obj).__array_ufunc__ is
                                np.ndarray.__array_ufunc__)):
                            # __array__ gets ignored
                            assert_(arr_imethod(obj) is arr, err_msg)
                        else:
                            assert_raises((TypeError, Coerced),
                                          arr_imethod, obj,
                                          err_msg=err_msg)

                    op_fn = getattr(operator, op, None)
                    if op_fn is None:
                        op_fn = getattr(operator, op + "_", None)
                    if op_fn is None:
                        op_fn = getattr(builtins, op)
                    assert_equal(op_fn(obj, arr), "forward", err_msg)
                    if not isinstance(obj, np.ndarray):
                        if binop_override_expected:
                            assert_equal(op_fn(arr, obj), "reverse", err_msg)
                        elif ufunc_override_expected:
                            assert_equal(op_fn(arr, obj)[0], "__array_ufunc__",
                                         err_msg)
                    if ufunc_override_expected:
                        assert_equal(ufunc(obj, arr)[0], "__array_ufunc__",
                                     err_msg)

        # No array priority, no array_ufunc -> nothing called
        check(make_obj(object), False, False, False)
        # Negative array priority, no array_ufunc -> nothing called
        # (has to be very negative, because scalar priority is -1000000.0)
        check(make_obj(object, array_priority=-2**30), False, False, False)
        # Positive array priority, no array_ufunc -> binops and iops only
        check(make_obj(object, array_priority=1), True, False, True)
        # ndarray ignores array_priority for ndarray subclasses
        check(make_obj(np.ndarray, array_priority=1), False, False, False,
              check_scalar=False)
        # Positive array_priority and array_ufunc -> array_ufunc only
        check(make_obj(object, array_priority=1,
                       array_ufunc=array_ufunc_impl), False, True, False)
        check(make_obj(np.ndarray, array_priority=1,
                       array_ufunc=array_ufunc_impl), False, True, False)
        # array_ufunc set to None -> defer binops only
        check(make_obj(object, array_ufunc=None), True, False, False)
        check(make_obj(np.ndarray, array_ufunc=None), True, False, False,
              check_scalar=False)

    @pytest.mark.parametrize("priority", [None, "runtime error"])
    def test_ufunc_binop_bad_array_priority(self, priority):
        # Mainly checks that this does not crash.  The second array has a lower
        # priority than -1 ("error value").  If the __radd__ actually exists,
        # bad things can happen (I think via the scalar paths).
        # In principle both of these can probably just be errors in the future.
        class BadPriority:
            @property
            def __array_priority__(self):
                if priority == "runtime error":
                    raise RuntimeError("RuntimeError in __array_priority__!")
                return priority

            def __radd__(self, other):
                return "result"

        class LowPriority(np.ndarray):
            __array_priority__ = -1000

        # Priority failure uses the same as scalars (smaller -1000).  So the
        # LowPriority wins with 'result' for each element (inner operation).
        res = np.arange(3).view(LowPriority) + BadPriority()
        assert res.shape == (3,)
        assert res[0] == 'result'

    @pytest.mark.parametrize("scalar", [
            np.longdouble(1), np.timedelta64(120, 'm')])
    @pytest.mark.parametrize("op", [operator.add, operator.xor])
    def test_scalar_binop_guarantees_ufunc(self, scalar, op):
        # Test that __array_ufunc__ will always cause ufunc use even when
        # we have to protect some other calls from recursing (see gh-26904).
        class SomeClass:
            def __array_ufunc__(self, ufunc, method, *inputs, **kw):
                return "result"

        assert SomeClass() + scalar == "result"
        assert scalar + SomeClass() == "result"

    def test_ufunc_override_normalize_signature(self):
        # gh-5674
        class SomeClass:
            def __array_ufunc__(self, ufunc, method, *inputs, **kw):
                return kw

        a = SomeClass()
        kw = np.add(a, [1])
        assert_('sig' not in kw and 'signature' not in kw)
        kw = np.add(a, [1], sig='ii->i')
        assert_('sig' not in kw and 'signature' in kw)
        assert_equal(kw['signature'], 'ii->i')
        kw = np.add(a, [1], signature='ii->i')
        assert_('sig' not in kw and 'signature' in kw)
        assert_equal(kw['signature'], 'ii->i')

    def test_array_ufunc_index(self):
        # Check that index is set appropriately, also if only an output
        # is passed on (latter is another regression tests for github bug 4753)
        # This also checks implicitly that 'out' is always a tuple.
        class CheckIndex:
            def __array_ufunc__(self, ufunc, method, *inputs, **kw):
                for i, a in enumerate(inputs):
                    if a is self:
                        return i
                # calls below mean we must be in an output.
                for j, a in enumerate(kw['out']):
                    if a is self:
                        return (j,)

        a = CheckIndex()
        dummy = np.arange(2.)
        # 1 input, 1 output
        assert_equal(np.sin(a), 0)
        assert_equal(np.sin(dummy, a), (0,))
        assert_equal(np.sin(dummy, out=a), (0,))
        assert_equal(np.sin(dummy, out=(a,)), (0,))
        assert_equal(np.sin(a, a), 0)
        assert_equal(np.sin(a, out=a), 0)
        assert_equal(np.sin(a, out=(a,)), 0)
        # 1 input, 2 outputs
        assert_equal(np.modf(dummy, a), (0,))
        assert_equal(np.modf(dummy, None, a), (1,))
        assert_equal(np.modf(dummy, dummy, a), (1,))
        assert_equal(np.modf(dummy, out=(a, None)), (0,))
        assert_equal(np.modf(dummy, out=(a, dummy)), (0,))
        assert_equal(np.modf(dummy, out=(None, a)), (1,))
        assert_equal(np.modf(dummy, out=(dummy, a)), (1,))
        assert_equal(np.modf(a, out=(dummy, a)), 0)
        with assert_raises(TypeError):
            # Out argument must be tuple, since there are multiple outputs
            np.modf(dummy, out=a)

        assert_raises(ValueError, np.modf, dummy, out=(a,))

        # 2 inputs, 1 output
        assert_equal(np.add(a, dummy), 0)
        assert_equal(np.add(dummy, a), 1)
        assert_equal(np.add(dummy, dummy, a), (0,))
        assert_equal(np.add(dummy, a, a), 1)
        assert_equal(np.add(dummy, dummy, out=a), (0,))
        assert_equal(np.add(dummy, dummy, out=(a,)), (0,))
        assert_equal(np.add(a, dummy, out=a), 0)

    def test_out_override(self):
        # regression test for github bug 4753
        class OutClass(np.ndarray):
            def __array_ufunc__(self, ufunc, method, *inputs, **kw):
                if 'out' in kw:
                    tmp_kw = kw.copy()
                    tmp_kw.pop('out')
                    func = getattr(ufunc, method)
                    kw['out'][0][...] = func(*inputs, **tmp_kw)

        A = np.array([0]).view(OutClass)
        B = np.array([5])
        C = np.array([6])
        np.multiply(C, B, A)
        assert_equal(A[0], 30)
        assert_(isinstance(A, OutClass))
        A[0] = 0
        np.multiply(C, B, out=A)
        assert_equal(A[0], 30)
        assert_(isinstance(A, OutClass))

    def test_pow_array_object_dtype(self):
        # test pow on arrays of object dtype
        class SomeClass:
            def __init__(self, num=None):
                self.num = num

            # want to ensure a fast pow path is not taken
            def __mul__(self, other):
                raise AssertionError('__mul__ should not be called')

            def __truediv__(self, other):
                raise AssertionError('__truediv__ should not be called')

            def __pow__(self, exp):
                return SomeClass(num=self.num ** exp)

            def __eq__(self, other):
                if isinstance(other, SomeClass):
                    return self.num == other.num

            __rpow__ = __pow__

        def pow_for(exp, arr):
            return np.array([x ** exp for x in arr])

        obj_arr = np.array([SomeClass(1), SomeClass(2), SomeClass(3)])

        assert_equal(obj_arr ** 0.5, pow_for(0.5, obj_arr))
        assert_equal(obj_arr ** 0, pow_for(0, obj_arr))
        assert_equal(obj_arr ** 1, pow_for(1, obj_arr))
        assert_equal(obj_arr ** -1, pow_for(-1, obj_arr))
        assert_equal(obj_arr ** 2, pow_for(2, obj_arr))

    def test_pos_array_ufunc_override(self):
        class A(np.ndarray):
            def __array_ufunc__(self, ufunc, method, *inputs, **kwargs):
                return getattr(ufunc, method)(*[i.view(np.ndarray) for
                                                i in inputs], **kwargs)
        tst = np.array('foo').view(A)
        with assert_raises(TypeError):
            +tst


class TestTemporaryElide:
    # elision is only triggered on relatively large arrays

    def test_extension_incref_elide(self):
        # test extension (e.g. cython) calling PyNumber_* slots without
        # increasing the reference counts
        #
        # def incref_elide(a):
        #    d = input.copy() # refcount 1
        #    return d, d + d # PyNumber_Add without increasing refcount
        from numpy._core._multiarray_tests import incref_elide
        d = np.ones(100000)
        orig, res = incref_elide(d)
        d + d
        # the return original should not be changed to an inplace operation
        assert_array_equal(orig, d)
        assert_array_equal(res, d + d)

    def test_extension_incref_elide_stack(self):
        # scanning if the refcount == 1 object is on the python stack to check
        # that we are called directly from python is flawed as object may still
        # be above the stack pointer and we have no access to the top of it
        #
        # def incref_elide_l(d):
        #    return l[4] + l[4] # PyNumber_Add without increasing refcount
        from numpy._core._multiarray_tests import incref_elide_l
        # padding with 1 makes sure the object on the stack is not overwritten
        l = [1, 1, 1, 1, np.ones(100000)]
        res = incref_elide_l(l)
        # the return original should not be changed to an inplace operation
        assert_array_equal(l[4], np.ones(100000))
        assert_array_equal(res, l[4] + l[4])

    def test_temporary_with_cast(self):
        # check that we don't elide into a temporary which would need casting
        d = np.ones(200000, dtype=np.int64)
        r = ((d + d) + np.array(2**222, dtype='O'))
        assert_equal(r.dtype, np.dtype('O'))

        r = ((d + d) / 2)
        assert_equal(r.dtype, np.dtype('f8'))

        r = np.true_divide((d + d), 2)
        assert_equal(r.dtype, np.dtype('f8'))

        r = ((d + d) / 2.)
        assert_equal(r.dtype, np.dtype('f8'))

        r = ((d + d) // 2)
        assert_equal(r.dtype, np.dtype(np.int64))

        # commutative elision into the astype result
        f = np.ones(100000, dtype=np.float32)
        assert_equal(((f + f) + f.astype(np.float64)).dtype, np.dtype('f8'))

        # no elision into lower type
        d = f.astype(np.float64)
        assert_equal(((f + f) + d).dtype, d.dtype)
        l = np.ones(100000, dtype=np.longdouble)
        assert_equal(((d + d) + l).dtype, l.dtype)

        # test unary abs with different output dtype
        for dt in (np.complex64, np.complex128, np.clongdouble):
            c = np.ones(100000, dtype=dt)
            r = abs(c * 2.0)
            assert_equal(r.dtype, np.dtype('f%d' % (c.itemsize // 2)))

    def test_elide_broadcast(self):
        # test no elision on broadcast to higher dimension
        # only triggers elision code path in debug mode as triggering it in
        # normal mode needs 256kb large matching dimension, so a lot of memory
        d = np.ones((2000, 1), dtype=int)
        b = np.ones((2000), dtype=bool)
        r = (1 - d) + b
        assert_equal(r, 1)
        assert_equal(r.shape, (2000, 2000))

    def test_elide_scalar(self):
        # check inplace op does not create ndarray from scalars
        a = np.bool()
        assert_(type(~(a & a)) is np.bool)

    def test_elide_scalar_readonly(self):
        # The imaginary part of a real array is readonly. This needs to go
        # through fast_scalar_power which is only called for powers of
        # +1, -1, 0, 0.5, and 2, so use 2. Also need valid refcount for
        # elision which can be gotten for the imaginary part of a real
        # array. Should not error.
        a = np.empty(100000, dtype=np.float64)
        a.imag ** 2

    def test_elide_readonly(self):
        # don't try to elide readonly temporaries
        r = np.asarray(np.broadcast_to(np.zeros(1), 100000).flat) * 0.0
        assert_equal(r, 0)

    def test_elide_updateifcopy(self):
        a = np.ones(2**20)[::2]
        b = a.flat.__array__() + 1
        del b
        assert_equal(a, 1)


class TestCAPI:
    def test_IsPythonScalar(self):
        from numpy._core._multiarray_tests import IsPythonScalar
        assert_(IsPythonScalar(b'foobar'))
        assert_(IsPythonScalar(1))
        assert_(IsPythonScalar(2**80))
        assert_(IsPythonScalar(2.))
        assert_(IsPythonScalar("a"))

    @pytest.mark.parametrize("converter",
             [_multiarray_tests.run_scalar_intp_converter,
              _multiarray_tests.run_scalar_intp_from_sequence])
    def test_intp_sequence_converters(self, converter):
        # Test simple values (-1 is special for error return paths)
        assert converter(10) == (10,)
        assert converter(-1) == (-1,)
        # A 0-D array looks a bit like a sequence but must take the integer
        # path:
        assert converter(np.array(123)) == (123,)
        # Test simple sequences (intp_from_sequence only supports length 1):
        assert converter((10,)) == (10,)
        assert converter(np.array([11])) == (11,)

    @pytest.mark.parametrize("converter",
             [_multiarray_tests.run_scalar_intp_converter,
              _multiarray_tests.run_scalar_intp_from_sequence])
    @pytest.mark.skipif(IS_PYPY and sys.implementation.version <= (7, 3, 8),
            reason="PyPy bug in error formatting")
    def test_intp_sequence_converters_errors(self, converter):
        with pytest.raises(TypeError,
                match="expected a sequence of integers or a single integer, "):
            converter(object())
        with pytest.raises(TypeError,
                match="expected a sequence of integers or a single integer, "
                      "got '32.0'"):
            converter(32.)
        with pytest.raises(TypeError,
                match="'float' object cannot be interpreted as an integer"):
            converter([32.])
        with pytest.raises(ValueError,
                match="Maximum allowed dimension"):
            # These converters currently convert overflows to a ValueError
            converter(2**64)


class TestSubscripting:
    def test_test_zero_rank(self):
        x = np.array([1, 2, 3])
        assert_(isinstance(x[0], np.int_))
        assert_(type(x[0, ...]) is np.ndarray)


class TestPickling:
    @pytest.mark.skipif(pickle.HIGHEST_PROTOCOL >= 5,
                        reason=('this tests the error messages when trying to'
                                'protocol 5 although it is not available'))
    def test_correct_protocol5_error_message(self):
        array = np.arange(10)

    def test_record_array_with_object_dtype(self):
        my_object = object()

        arr_with_object = np.array(
                [(my_object, 1, 2.0)],
                dtype=[('a', object), ('b', int), ('c', float)])
        arr_without_object = np.array(
                [('xxx', 1, 2.0)],
                dtype=[('a', str), ('b', int), ('c', float)])

        for proto in range(2, pickle.HIGHEST_PROTOCOL + 1):
            depickled_arr_with_object = pickle.loads(
                    pickle.dumps(arr_with_object, protocol=proto))
            depickled_arr_without_object = pickle.loads(
                    pickle.dumps(arr_without_object, protocol=proto))

            assert_equal(arr_with_object.dtype,
                         depickled_arr_with_object.dtype)
            assert_equal(arr_without_object.dtype,
                         depickled_arr_without_object.dtype)

    @pytest.mark.skipif(pickle.HIGHEST_PROTOCOL < 5,
                        reason="requires pickle protocol 5")
    def test_f_contiguous_array(self):
        f_contiguous_array = np.array([[1, 2, 3], [4, 5, 6]], order='F')
        buffers = []

        # When using pickle protocol 5, Fortran-contiguous arrays can be
        # serialized using out-of-band buffers
        bytes_string = pickle.dumps(f_contiguous_array, protocol=5,
                                    buffer_callback=buffers.append)

        assert len(buffers) > 0

        depickled_f_contiguous_array = pickle.loads(bytes_string,
                                                    buffers=buffers)

        assert_equal(f_contiguous_array, depickled_f_contiguous_array)

    @pytest.mark.skipif(pickle.HIGHEST_PROTOCOL < 5, reason="requires pickle protocol 5")
    @pytest.mark.parametrize('transposed_contiguous_array',
        [np.random.default_rng(42).random((2, 3, 4)).transpose((1, 0, 2)),
         np.random.default_rng(42).random((2, 3, 4, 5)).transpose((1, 3, 0, 2))] +
        [np.random.default_rng(42).random(np.arange(2, 7)).transpose(np.random.permutation(5)) for _ in range(3)])
    def test_transposed_contiguous_array(self, transposed_contiguous_array):
        buffers = []
        # When using pickle protocol 5, arrays which can be transposed to c_contiguous
        # can be serialized using out-of-band buffers
        bytes_string = pickle.dumps(transposed_contiguous_array, protocol=5,
                                    buffer_callback=buffers.append)

        assert len(buffers) > 0

        depickled_transposed_contiguous_array = pickle.loads(bytes_string,
                                                    buffers=buffers)

        assert_equal(transposed_contiguous_array, depickled_transposed_contiguous_array)

    @pytest.mark.skipif(pickle.HIGHEST_PROTOCOL < 5, reason="requires pickle protocol 5")
    def test_load_legacy_pkl_protocol5(self):
        # legacy byte strs are dumped in 2.2.1
        c_contiguous_dumped = b'\x80\x05\x95\x90\x00\x00\x00\x00\x00\x00\x00\x8c\x13numpy._core.numeric\x94\x8c\x0b_frombuffer\x94\x93\x94(\x96\x18\x00\x00\x00\x00\x00\x00\x00\x00\x01\x02\x03\x04\x05\x06\x07\x08\t\n\x0b\x0c\r\x0e\x0f\x10\x11\x12\x13\x14\x15\x16\x17\x94\x8c\x05numpy\x94\x8c\x05dtype\x94\x93\x94\x8c\x02u1\x94\x89\x88\x87\x94R\x94(K\x03\x8c\x01|\x94NNNJ\xff\xff\xff\xffJ\xff\xff\xff\xffK\x00t\x94bK\x03K\x04K\x02\x87\x94\x8c\x01C\x94t\x94R\x94.'  # noqa: E501
        f_contiguous_dumped = b'\x80\x05\x95\x90\x00\x00\x00\x00\x00\x00\x00\x8c\x13numpy._core.numeric\x94\x8c\x0b_frombuffer\x94\x93\x94(\x96\x18\x00\x00\x00\x00\x00\x00\x00\x00\x01\x02\x03\x04\x05\x06\x07\x08\t\n\x0b\x0c\r\x0e\x0f\x10\x11\x12\x13\x14\x15\x16\x17\x94\x8c\x05numpy\x94\x8c\x05dtype\x94\x93\x94\x8c\x02u1\x94\x89\x88\x87\x94R\x94(K\x03\x8c\x01|\x94NNNJ\xff\xff\xff\xffJ\xff\xff\xff\xffK\x00t\x94bK\x03K\x04K\x02\x87\x94\x8c\x01F\x94t\x94R\x94.'   # noqa: E501
        transposed_contiguous_dumped = b'\x80\x05\x95\xa5\x00\x00\x00\x00\x00\x00\x00\x8c\x16numpy._core.multiarray\x94\x8c\x0c_reconstruct\x94\x93\x94\x8c\x05numpy\x94\x8c\x07ndarray\x94\x93\x94K\x00\x85\x94C\x01b\x94\x87\x94R\x94(K\x01K\x04K\x03K\x02\x87\x94h\x03\x8c\x05dtype\x94\x93\x94\x8c\x02u1\x94\x89\x88\x87\x94R\x94(K\x03\x8c\x01|\x94NNNJ\xff\xff\xff\xffJ\xff\xff\xff\xffK\x00t\x94b\x89C\x18\x00\x01\x08\t\x10\x11\x02\x03\n\x0b\x12\x13\x04\x05\x0c\r\x14\x15\x06\x07\x0e\x0f\x16\x17\x94t\x94b.'   # noqa: E501
        no_contiguous_dumped = b'\x80\x05\x95\x91\x00\x00\x00\x00\x00\x00\x00\x8c\x16numpy._core.multiarray\x94\x8c\x0c_reconstruct\x94\x93\x94\x8c\x05numpy\x94\x8c\x07ndarray\x94\x93\x94K\x00\x85\x94C\x01b\x94\x87\x94R\x94(K\x01K\x03K\x02\x86\x94h\x03\x8c\x05dtype\x94\x93\x94\x8c\x02u1\x94\x89\x88\x87\x94R\x94(K\x03\x8c\x01|\x94NNNJ\xff\xff\xff\xffJ\xff\xff\xff\xffK\x00t\x94b\x89C\x06\x00\x01\x04\x05\x08\t\x94t\x94b.'  # noqa: E501
        x = np.arange(24, dtype='uint8').reshape(3, 4, 2)
        assert_equal(x, pickle.loads(c_contiguous_dumped))
        x = np.arange(24, dtype='uint8').reshape(3, 4, 2, order='F')
        assert_equal(x, pickle.loads(f_contiguous_dumped))
        x = np.arange(24, dtype='uint8').reshape(3, 4, 2).transpose((1, 0, 2))
        assert_equal(x, pickle.loads(transposed_contiguous_dumped))
        x = np.arange(12, dtype='uint8').reshape(3, 4)[:, :2]
        assert_equal(x, pickle.loads(no_contiguous_dumped))

    def test_non_contiguous_array(self):
        non_contiguous_array = np.arange(12).reshape(3, 4)[:, :2]
        assert not non_contiguous_array.flags.c_contiguous
        assert not non_contiguous_array.flags.f_contiguous

        # make sure non-contiguous arrays can be pickled-depickled
        # using any protocol
        buffers = []
        for proto in range(2, pickle.HIGHEST_PROTOCOL + 1):
            depickled_non_contiguous_array = pickle.loads(
                    pickle.dumps(non_contiguous_array, protocol=proto,
                                 buffer_callback=buffers.append if proto >= 5 else None))

            assert_equal(len(buffers), 0)
            assert_equal(non_contiguous_array, depickled_non_contiguous_array)

    def test_roundtrip(self):
        for proto in range(2, pickle.HIGHEST_PROTOCOL + 1):
            carray = np.array([[2, 9], [7, 0], [3, 8]])
            DATA = [
                carray,
                np.transpose(carray),
                np.array([('xxx', 1, 2.0)], dtype=[('a', (str, 3)), ('b', int),
                                                   ('c', float)])
            ]

            refs = [weakref.ref(a) for a in DATA]
            for a in DATA:
                assert_equal(
                        a, pickle.loads(pickle.dumps(a, protocol=proto)),
                        err_msg=f"{a!r}")
            del a, DATA, carray
            break_cycles()
            # check for reference leaks (gh-12793)
            for ref in refs:
                assert ref() is None

    def _loads(self, obj):
        return pickle.loads(obj, encoding='latin1')

    # version 0 pickles, using protocol=2 to pickle
    # version 0 doesn't have a version field
    def test_version0_int8(self):
        s = b"\x80\x02cnumpy.core._internal\n_reconstruct\nq\x01cnumpy\nndarray\nq\x02K\x00\x85U\x01b\x87Rq\x03(K\x04\x85cnumpy\ndtype\nq\x04U\x02i1K\x00K\x01\x87Rq\x05(U\x01|NNJ\xff\xff\xff\xffJ\xff\xff\xff\xfftb\x89U\x04\x01\x02\x03\x04tb."
        a = np.array([1, 2, 3, 4], dtype=np.int8)
        p = self._loads(s)
        assert_equal(a, p)

    def test_version0_float32(self):
        s = b"\x80\x02cnumpy.core._internal\n_reconstruct\nq\x01cnumpy\nndarray\nq\x02K\x00\x85U\x01b\x87Rq\x03(K\x04\x85cnumpy\ndtype\nq\x04U\x02f4K\x00K\x01\x87Rq\x05(U\x01<NNJ\xff\xff\xff\xffJ\xff\xff\xff\xfftb\x89U\x10\x00\x00\x80?\x00\x00\x00@\x00\x00@@\x00\x00\x80@tb."
        a = np.array([1.0, 2.0, 3.0, 4.0], dtype=np.float32)
        p = self._loads(s)
        assert_equal(a, p)

    def test_version0_object(self):
        s = b"\x80\x02cnumpy.core._internal\n_reconstruct\nq\x01cnumpy\nndarray\nq\x02K\x00\x85U\x01b\x87Rq\x03(K\x02\x85cnumpy\ndtype\nq\x04U\x02O8K\x00K\x01\x87Rq\x05(U\x01|NNJ\xff\xff\xff\xffJ\xff\xff\xff\xfftb\x89]q\x06(}q\x07U\x01aK\x01s}q\x08U\x01bK\x02setb."
        a = np.array([{'a': 1}, {'b': 2}])
        p = self._loads(s)
        assert_equal(a, p)

    # version 1 pickles, using protocol=2 to pickle
    def test_version1_int8(self):
        s = b"\x80\x02cnumpy.core._internal\n_reconstruct\nq\x01cnumpy\nndarray\nq\x02K\x00\x85U\x01b\x87Rq\x03(K\x01K\x04\x85cnumpy\ndtype\nq\x04U\x02i1K\x00K\x01\x87Rq\x05(K\x01U\x01|NNJ\xff\xff\xff\xffJ\xff\xff\xff\xfftb\x89U\x04\x01\x02\x03\x04tb."
        a = np.array([1, 2, 3, 4], dtype=np.int8)
        p = self._loads(s)
        assert_equal(a, p)

    def test_version1_float32(self):
        s = b"\x80\x02cnumpy.core._internal\n_reconstruct\nq\x01cnumpy\nndarray\nq\x02K\x00\x85U\x01b\x87Rq\x03(K\x01K\x04\x85cnumpy\ndtype\nq\x04U\x02f4K\x00K\x01\x87Rq\x05(K\x01U\x01<NNJ\xff\xff\xff\xffJ\xff\xff\xff\xfftb\x89U\x10\x00\x00\x80?\x00\x00\x00@\x00\x00@@\x00\x00\x80@tb."
        a = np.array([1.0, 2.0, 3.0, 4.0], dtype=np.float32)
        p = self._loads(s)
        assert_equal(a, p)

    def test_version1_object(self):
        s = b"\x80\x02cnumpy.core._internal\n_reconstruct\nq\x01cnumpy\nndarray\nq\x02K\x00\x85U\x01b\x87Rq\x03(K\x01K\x02\x85cnumpy\ndtype\nq\x04U\x02O8K\x00K\x01\x87Rq\x05(K\x01U\x01|NNJ\xff\xff\xff\xffJ\xff\xff\xff\xfftb\x89]q\x06(}q\x07U\x01aK\x01s}q\x08U\x01bK\x02setb."
        a = np.array([{'a': 1}, {'b': 2}])
        p = self._loads(s)
        assert_equal(a, p)

    def test_subarray_int_shape(self):
        s = b"cnumpy.core.multiarray\n_reconstruct\np0\n(cnumpy\nndarray\np1\n(I0\ntp2\nS'b'\np3\ntp4\nRp5\n(I1\n(I1\ntp6\ncnumpy\ndtype\np7\n(S'V6'\np8\nI0\nI1\ntp9\nRp10\n(I3\nS'|'\np11\nN(S'a'\np12\ng3\ntp13\n(dp14\ng12\n(g7\n(S'V4'\np15\nI0\nI1\ntp16\nRp17\n(I3\nS'|'\np18\n(g7\n(S'i1'\np19\nI0\nI1\ntp20\nRp21\n(I3\nS'|'\np22\nNNNI-1\nI-1\nI0\ntp23\nb(I2\nI2\ntp24\ntp25\nNNI4\nI1\nI0\ntp26\nbI0\ntp27\nsg3\n(g7\n(S'V2'\np28\nI0\nI1\ntp29\nRp30\n(I3\nS'|'\np31\n(g21\nI2\ntp32\nNNI2\nI1\nI0\ntp33\nbI4\ntp34\nsI6\nI1\nI0\ntp35\nbI00\nS'\\x01\\x01\\x01\\x01\\x01\\x02'\np36\ntp37\nb."
        a = np.array([(1, (1, 2))], dtype=[('a', 'i1', (2, 2)), ('b', 'i1', 2)])
        p = self._loads(s)
        assert_equal(a, p)

    def test_datetime64_byteorder(self):
        original = np.array([['2015-02-24T00:00:00.000000000']], dtype='datetime64[ns]')

        original_byte_reversed = original.copy(order='K')
        original_byte_reversed.dtype = original_byte_reversed.dtype.newbyteorder('S')
        original_byte_reversed.byteswap(inplace=True)

        new = pickle.loads(pickle.dumps(original_byte_reversed))

        assert_equal(original.dtype, new.dtype)


class TestFancyIndexing:
    def test_list(self):
        x = np.ones((1, 1))
        x[:, [0]] = 2.0
        assert_array_equal(x, np.array([[2.0]]))

        x = np.ones((1, 1, 1))
        x[:, :, [0]] = 2.0
        assert_array_equal(x, np.array([[[2.0]]]))

    def test_tuple(self):
        x = np.ones((1, 1))
        x[:, (0,)] = 2.0
        assert_array_equal(x, np.array([[2.0]]))
        x = np.ones((1, 1, 1))
        x[:, :, (0,)] = 2.0
        assert_array_equal(x, np.array([[[2.0]]]))

    def test_mask(self):
        x = np.array([1, 2, 3, 4])
        m = np.array([0, 1, 0, 0], bool)
        assert_array_equal(x[m], np.array([2]))

    def test_mask2(self):
        x = np.array([[1, 2, 3, 4], [5, 6, 7, 8]])
        m = np.array([0, 1], bool)
        m2 = np.array([[0, 1, 0, 0], [1, 0, 0, 0]], bool)
        m3 = np.array([[0, 1, 0, 0], [0, 0, 0, 0]], bool)
        assert_array_equal(x[m], np.array([[5, 6, 7, 8]]))
        assert_array_equal(x[m2], np.array([2, 5]))
        assert_array_equal(x[m3], np.array([2]))

    def test_assign_mask(self):
        x = np.array([1, 2, 3, 4])
        m = np.array([0, 1, 0, 0], bool)
        x[m] = 5
        assert_array_equal(x, np.array([1, 5, 3, 4]))

    def test_assign_mask2(self):
        xorig = np.array([[1, 2, 3, 4], [5, 6, 7, 8]])
        m = np.array([0, 1], bool)
        m2 = np.array([[0, 1, 0, 0], [1, 0, 0, 0]], bool)
        m3 = np.array([[0, 1, 0, 0], [0, 0, 0, 0]], bool)
        x = xorig.copy()
        x[m] = 10
        assert_array_equal(x, np.array([[1, 2, 3, 4], [10, 10, 10, 10]]))
        x = xorig.copy()
        x[m2] = 10
        assert_array_equal(x, np.array([[1, 10, 3, 4], [10, 6, 7, 8]]))
        x = xorig.copy()
        x[m3] = 10
        assert_array_equal(x, np.array([[1, 10, 3, 4], [5, 6, 7, 8]]))


class TestStringCompare:
    def test_string(self):
        g1 = np.array(["This", "is", "example"])
        g2 = np.array(["This", "was", "example"])
        assert_array_equal(g1 == g2, [g1[i] == g2[i] for i in [0, 1, 2]])
        assert_array_equal(g1 != g2, [g1[i] != g2[i] for i in [0, 1, 2]])
        assert_array_equal(g1 <= g2, [g1[i] <= g2[i] for i in [0, 1, 2]])
        assert_array_equal(g1 >= g2, [g1[i] >= g2[i] for i in [0, 1, 2]])
        assert_array_equal(g1 < g2, [g1[i] < g2[i] for i in [0, 1, 2]])
        assert_array_equal(g1 > g2, [g1[i] > g2[i] for i in [0, 1, 2]])

    def test_mixed(self):
        g1 = np.array(["spam", "spa", "spammer", "and eggs"])
        g2 = "spam"
        assert_array_equal(g1 == g2, [x == g2 for x in g1])
        assert_array_equal(g1 != g2, [x != g2 for x in g1])
        assert_array_equal(g1 < g2, [x < g2 for x in g1])
        assert_array_equal(g1 > g2, [x > g2 for x in g1])
        assert_array_equal(g1 <= g2, [x <= g2 for x in g1])
        assert_array_equal(g1 >= g2, [x >= g2 for x in g1])

    def test_unicode(self):
        g1 = np.array(["This", "is", "example"])
        g2 = np.array(["This", "was", "example"])
        assert_array_equal(g1 == g2, [g1[i] == g2[i] for i in [0, 1, 2]])
        assert_array_equal(g1 != g2, [g1[i] != g2[i] for i in [0, 1, 2]])
        assert_array_equal(g1 <= g2, [g1[i] <= g2[i] for i in [0, 1, 2]])
        assert_array_equal(g1 >= g2, [g1[i] >= g2[i] for i in [0, 1, 2]])
        assert_array_equal(g1 < g2,  [g1[i] < g2[i] for i in [0, 1, 2]])
        assert_array_equal(g1 > g2,  [g1[i] > g2[i] for i in [0, 1, 2]])

class TestArgmaxArgminCommon:

    sizes = [(), (3,), (3, 2), (2, 3),
             (3, 3), (2, 3, 4), (4, 3, 2),
             (1, 2, 3, 4), (2, 3, 4, 1),
             (3, 4, 1, 2), (4, 1, 2, 3),
             (64,), (128,), (256,)]

    @pytest.mark.parametrize("size, axis", itertools.chain(*[[(size, axis)
        for axis in list(range(-len(size), len(size))) + [None]]
        for size in sizes]))
    @pytest.mark.parametrize('method', [np.argmax, np.argmin])
    def test_np_argmin_argmax_keepdims(self, size, axis, method):

        arr = np.random.normal(size=size)

        # contiguous arrays
        if axis is None:
            new_shape = [1 for _ in range(len(size))]
        else:
            new_shape = list(size)
            new_shape[axis] = 1
        new_shape = tuple(new_shape)

        _res_orig = method(arr, axis=axis)
        res_orig = _res_orig.reshape(new_shape)
        res = method(arr, axis=axis, keepdims=True)
        assert_equal(res, res_orig)
        assert_(res.shape == new_shape)
        outarray = np.empty(res.shape, dtype=res.dtype)
        res1 = method(arr, axis=axis, out=outarray,
                            keepdims=True)
        assert_(res1 is outarray)
        assert_equal(res, outarray)

        if len(size) > 0:
            wrong_shape = list(new_shape)
            if axis is not None:
                wrong_shape[axis] = 2
            else:
                wrong_shape[0] = 2
            wrong_outarray = np.empty(wrong_shape, dtype=res.dtype)
            with pytest.raises(ValueError):
                method(arr.T, axis=axis,
                        out=wrong_outarray, keepdims=True)

        # non-contiguous arrays
        if axis is None:
            new_shape = [1 for _ in range(len(size))]
        else:
            new_shape = list(size)[::-1]
            new_shape[axis] = 1
        new_shape = tuple(new_shape)

        _res_orig = method(arr.T, axis=axis)
        res_orig = _res_orig.reshape(new_shape)
        res = method(arr.T, axis=axis, keepdims=True)
        assert_equal(res, res_orig)
        assert_(res.shape == new_shape)
        outarray = np.empty(new_shape[::-1], dtype=res.dtype)
        outarray = outarray.T
        res1 = method(arr.T, axis=axis, out=outarray,
                            keepdims=True)
        assert_(res1 is outarray)
        assert_equal(res, outarray)

        if len(size) > 0:
            # one dimension lesser for non-zero sized
            # array should raise an error
            with pytest.raises(ValueError):
                method(arr[0], axis=axis,
                        out=outarray, keepdims=True)

        if len(size) > 0:
            wrong_shape = list(new_shape)
            if axis is not None:
                wrong_shape[axis] = 2
            else:
                wrong_shape[0] = 2
            wrong_outarray = np.empty(wrong_shape, dtype=res.dtype)
            with pytest.raises(ValueError):
                method(arr.T, axis=axis,
                        out=wrong_outarray, keepdims=True)

    @pytest.mark.parametrize('method', ['max', 'min'])
    def test_all(self, method):
        a = np.random.normal(0, 1, (4, 5, 6, 7, 8))
        arg_method = getattr(a, 'arg' + method)
        val_method = getattr(a, method)
        for i in range(a.ndim):
            a_maxmin = val_method(i)
            aarg_maxmin = arg_method(i)
            axes = list(range(a.ndim))
            axes.remove(i)
            assert_(np.all(a_maxmin == aarg_maxmin.choose(
                                        *a.transpose(i, *axes))))

    @pytest.mark.parametrize('method', ['argmax', 'argmin'])
    def test_output_shape(self, method):
        # see also gh-616
        a = np.ones((10, 5))
        arg_method = getattr(a, method)
        # Check some simple shape mismatches
        out = np.ones(11, dtype=np.int_)
        assert_raises(ValueError, arg_method, -1, out)

        out = np.ones((2, 5), dtype=np.int_)
        assert_raises(ValueError, arg_method, -1, out)

        # these could be relaxed possibly (used to allow even the previous)
        out = np.ones((1, 10), dtype=np.int_)
        assert_raises(ValueError, arg_method, -1, out)

        out = np.ones(10, dtype=np.int_)
        arg_method(-1, out=out)
        assert_equal(out, arg_method(-1))

    @pytest.mark.parametrize('ndim', [0, 1])
    @pytest.mark.parametrize('method', ['argmax', 'argmin'])
    def test_ret_is_out(self, ndim, method):
        a = np.ones((4,) + (256,) * ndim)
        arg_method = getattr(a, method)
        out = np.empty((256,) * ndim, dtype=np.intp)
        ret = arg_method(axis=0, out=out)
        assert ret is out

    @pytest.mark.parametrize('np_array, method, idx, val',
        [(np.zeros, 'argmax', 5942, "as"),
         (np.ones, 'argmin', 6001, "0")])
    def test_unicode(self, np_array, method, idx, val):
        d = np_array(6031, dtype='<U9')
        arg_method = getattr(d, method)
        d[idx] = val
        assert_equal(arg_method(), idx)

    @pytest.mark.parametrize('arr_method, np_method',
        [('argmax', np.argmax),
         ('argmin', np.argmin)])
    def test_np_vs_ndarray(self, arr_method, np_method):
        # make sure both ndarray.argmax/argmin and
        # numpy.argmax/argmin support out/axis args
        a = np.random.normal(size=(2, 3))
        arg_method = getattr(a, arr_method)

        # check positional args
        out1 = np.zeros(2, dtype=int)
        out2 = np.zeros(2, dtype=int)
        assert_equal(arg_method(1, out1), np_method(a, 1, out2))
        assert_equal(out1, out2)

        # check keyword args
        out1 = np.zeros(3, dtype=int)
        out2 = np.zeros(3, dtype=int)
        assert_equal(arg_method(out=out1, axis=0),
                     np_method(a, out=out2, axis=0))
        assert_equal(out1, out2)

    @pytest.mark.leaks_references(reason="replaces None with NULL.")
    @pytest.mark.parametrize('method, vals',
        [('argmax', (10, 30)),
         ('argmin', (30, 10))])
    def test_object_with_NULLs(self, method, vals):
        # See gh-6032
        a = np.empty(4, dtype='O')
        arg_method = getattr(a, method)
        ctypes.memset(a.ctypes.data, 0, a.nbytes)
        assert_equal(arg_method(), 0)
        a[3] = vals[0]
        assert_equal(arg_method(), 3)
        a[1] = vals[1]
        assert_equal(arg_method(), 1)

class TestArgmax:
    usg_data = [
        ([1, 1, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0], 0),
        ([3, 3, 3, 3,  2,  2,  2,  2], 0),
        ([0, 1, 2, 3,  4,  5,  6,  7], 7),
        ([7, 6, 5, 4,  3,  2,  1,  0], 0)
    ]
    sg_data = usg_data + [
        ([1, 2, 3, 4, -4, -3, -2, -1], 3),
        ([1, 2, 3, 4, -1, -2, -3, -4], 3)
    ]
    darr = [(np.array(d[0], dtype=t), d[1]) for d, t in (
        itertools.product(usg_data, (
            np.uint8, np.uint16, np.uint32, np.uint64
        ))
    )]
    darr = darr + [(np.array(d[0], dtype=t), d[1]) for d, t in (
        itertools.product(sg_data, (
            np.int8, np.int16, np.int32, np.int64, np.float32, np.float64
        ))
    )]
    darr = darr + [(np.array(d[0], dtype=t), d[1]) for d, t in (
        itertools.product((
            ([0, 1, 2, 3, np.nan], 4),
            ([0, 1, 2, np.nan, 3], 3),
            ([np.nan, 0, 1, 2, 3], 0),
            ([np.nan, 0, np.nan, 2, 3], 0),
            # To hit the tail of SIMD multi-level(x4, x1) inner loops
            # on variant SIMD widths
            ([1] * (2 * 5 - 1) + [np.nan], 2 * 5 - 1),
            ([1] * (4 * 5 - 1) + [np.nan], 4 * 5 - 1),
            ([1] * (8 * 5 - 1) + [np.nan], 8 * 5 - 1),
            ([1] * (16 * 5 - 1) + [np.nan], 16 * 5 - 1),
            ([1] * (32 * 5 - 1) + [np.nan], 32 * 5 - 1)
        ), (
            np.float32, np.float64
        ))
    )]
    nan_arr = darr + [
        ([0, 1, 2, 3, complex(0, np.nan)], 4),
        ([0, 1, 2, 3, complex(np.nan, 0)], 4),
        ([0, 1, 2, complex(np.nan, 0), 3], 3),
        ([0, 1, 2, complex(0, np.nan), 3], 3),
        ([complex(0, np.nan), 0, 1, 2, 3], 0),
        ([complex(np.nan, np.nan), 0, 1, 2, 3], 0),
        ([complex(np.nan, 0), complex(np.nan, 2), complex(np.nan, 1)], 0),
        ([complex(np.nan, np.nan), complex(np.nan, 2), complex(np.nan, 1)], 0),
        ([complex(np.nan, 0), complex(np.nan, 2), complex(np.nan, np.nan)], 0),

        ([complex(0, 0), complex(0, 2), complex(0, 1)], 1),
        ([complex(1, 0), complex(0, 2), complex(0, 1)], 0),
        ([complex(1, 0), complex(0, 2), complex(1, 1)], 2),

        ([np.datetime64('1923-04-14T12:43:12'),
          np.datetime64('1994-06-21T14:43:15'),
          np.datetime64('2001-10-15T04:10:32'),
          np.datetime64('1995-11-25T16:02:16'),
          np.datetime64('2005-01-04T03:14:12'),
          np.datetime64('2041-12-03T14:05:03')], 5),
        ([np.datetime64('1935-09-14T04:40:11'),
          np.datetime64('1949-10-12T12:32:11'),
          np.datetime64('2010-01-03T05:14:12'),
          np.datetime64('2015-11-20T12:20:59'),
          np.datetime64('1932-09-23T10:10:13'),
          np.datetime64('2014-10-10T03:50:30')], 3),
        # Assorted tests with NaTs
        ([np.datetime64('NaT'),
          np.datetime64('NaT'),
          np.datetime64('2010-01-03T05:14:12'),
          np.datetime64('NaT'),
          np.datetime64('2015-09-23T10:10:13'),
          np.datetime64('1932-10-10T03:50:30')], 0),
        ([np.datetime64('2059-03-14T12:43:12'),
          np.datetime64('1996-09-21T14:43:15'),
          np.datetime64('NaT'),
          np.datetime64('2022-12-25T16:02:16'),
          np.datetime64('1963-10-04T03:14:12'),
          np.datetime64('2013-05-08T18:15:23')], 2),
        ([np.timedelta64(2, 's'),
          np.timedelta64(1, 's'),
          np.timedelta64('NaT', 's'),
          np.timedelta64(3, 's')], 2),
        ([np.timedelta64('NaT', 's')] * 3, 0),

        ([timedelta(days=5, seconds=14), timedelta(days=2, seconds=35),
          timedelta(days=-1, seconds=23)], 0),
        ([timedelta(days=1, seconds=43), timedelta(days=10, seconds=5),
          timedelta(days=5, seconds=14)], 1),
        ([timedelta(days=10, seconds=24), timedelta(days=10, seconds=5),
          timedelta(days=10, seconds=43)], 2),

        ([False, False, False, False, True], 4),
        ([False, False, False, True, False], 3),
        ([True, False, False, False, False], 0),
        ([True, False, True, False, False], 0),
    ]

    @pytest.mark.parametrize('data', nan_arr)
    def test_combinations(self, data):
        arr, pos = data
        with suppress_warnings() as sup:
            sup.filter(RuntimeWarning,
                        "invalid value encountered in reduce")
            val = np.max(arr)

        assert_equal(np.argmax(arr), pos, err_msg=f"{arr!r}")
        assert_equal(arr[np.argmax(arr)], val, err_msg=f"{arr!r}")

        # add padding to test SIMD loops
        rarr = np.repeat(arr, 129)
        rpos = pos * 129
        assert_equal(np.argmax(rarr), rpos, err_msg=f"{rarr!r}")
        assert_equal(rarr[np.argmax(rarr)], val, err_msg=f"{rarr!r}")

        padd = np.repeat(np.min(arr), 513)
        rarr = np.concatenate((arr, padd))
        rpos = pos
        assert_equal(np.argmax(rarr), rpos, err_msg=f"{rarr!r}")
        assert_equal(rarr[np.argmax(rarr)], val, err_msg=f"{rarr!r}")

    def test_maximum_signed_integers(self):

        a = np.array([1, 2**7 - 1, -2**7], dtype=np.int8)
        assert_equal(np.argmax(a), 1)
        a = a.repeat(129)
        assert_equal(np.argmax(a), 129)

        a = np.array([1, 2**15 - 1, -2**15], dtype=np.int16)
        assert_equal(np.argmax(a), 1)
        a = a.repeat(129)
        assert_equal(np.argmax(a), 129)

        a = np.array([1, 2**31 - 1, -2**31], dtype=np.int32)
        assert_equal(np.argmax(a), 1)
        a = a.repeat(129)
        assert_equal(np.argmax(a), 129)

        a = np.array([1, 2**63 - 1, -2**63], dtype=np.int64)
        assert_equal(np.argmax(a), 1)
        a = a.repeat(129)
        assert_equal(np.argmax(a), 129)

class TestArgmin:
    usg_data = [
        ([1, 1, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0], 8),
        ([3, 3, 3, 3,  2,  2,  2,  2], 4),
        ([0, 1, 2, 3,  4,  5,  6,  7], 0),
        ([7, 6, 5, 4,  3,  2,  1,  0], 7)
    ]
    sg_data = usg_data + [
        ([1, 2, 3, 4, -4, -3, -2, -1], 4),
        ([1, 2, 3, 4, -1, -2, -3, -4], 7)
    ]
    darr = [(np.array(d[0], dtype=t), d[1]) for d, t in (
        itertools.product(usg_data, (
            np.uint8, np.uint16, np.uint32, np.uint64
        ))
    )]
    darr = darr + [(np.array(d[0], dtype=t), d[1]) for d, t in (
        itertools.product(sg_data, (
            np.int8, np.int16, np.int32, np.int64, np.float32, np.float64
        ))
    )]
    darr = darr + [(np.array(d[0], dtype=t), d[1]) for d, t in (
        itertools.product((
            ([0, 1, 2, 3, np.nan], 4),
            ([0, 1, 2, np.nan, 3], 3),
            ([np.nan, 0, 1, 2, 3], 0),
            ([np.nan, 0, np.nan, 2, 3], 0),
            # To hit the tail of SIMD multi-level(x4, x1) inner loops
            # on variant SIMD widths
            ([1] * (2 * 5 - 1) + [np.nan], 2 * 5 - 1),
            ([1] * (4 * 5 - 1) + [np.nan], 4 * 5 - 1),
            ([1] * (8 * 5 - 1) + [np.nan], 8 * 5 - 1),
            ([1] * (16 * 5 - 1) + [np.nan], 16 * 5 - 1),
            ([1] * (32 * 5 - 1) + [np.nan], 32 * 5 - 1)
        ), (
            np.float32, np.float64
        ))
    )]
    nan_arr = darr + [
        ([0, 1, 2, 3, complex(0, np.nan)], 4),
        ([0, 1, 2, 3, complex(np.nan, 0)], 4),
        ([0, 1, 2, complex(np.nan, 0), 3], 3),
        ([0, 1, 2, complex(0, np.nan), 3], 3),
        ([complex(0, np.nan), 0, 1, 2, 3], 0),
        ([complex(np.nan, np.nan), 0, 1, 2, 3], 0),
        ([complex(np.nan, 0), complex(np.nan, 2), complex(np.nan, 1)], 0),
        ([complex(np.nan, np.nan), complex(np.nan, 2), complex(np.nan, 1)], 0),
        ([complex(np.nan, 0), complex(np.nan, 2), complex(np.nan, np.nan)], 0),

        ([complex(0, 0), complex(0, 2), complex(0, 1)], 0),
        ([complex(1, 0), complex(0, 2), complex(0, 1)], 2),
        ([complex(1, 0), complex(0, 2), complex(1, 1)], 1),

        ([np.datetime64('1923-04-14T12:43:12'),
          np.datetime64('1994-06-21T14:43:15'),
          np.datetime64('2001-10-15T04:10:32'),
          np.datetime64('1995-11-25T16:02:16'),
          np.datetime64('2005-01-04T03:14:12'),
          np.datetime64('2041-12-03T14:05:03')], 0),
        ([np.datetime64('1935-09-14T04:40:11'),
          np.datetime64('1949-10-12T12:32:11'),
          np.datetime64('2010-01-03T05:14:12'),
          np.datetime64('2014-11-20T12:20:59'),
          np.datetime64('2015-09-23T10:10:13'),
          np.datetime64('1932-10-10T03:50:30')], 5),
        # Assorted tests with NaTs
        ([np.datetime64('NaT'),
          np.datetime64('NaT'),
          np.datetime64('2010-01-03T05:14:12'),
          np.datetime64('NaT'),
          np.datetime64('2015-09-23T10:10:13'),
          np.datetime64('1932-10-10T03:50:30')], 0),
        ([np.datetime64('2059-03-14T12:43:12'),
          np.datetime64('1996-09-21T14:43:15'),
          np.datetime64('NaT'),
          np.datetime64('2022-12-25T16:02:16'),
          np.datetime64('1963-10-04T03:14:12'),
          np.datetime64('2013-05-08T18:15:23')], 2),
        ([np.timedelta64(2, 's'),
          np.timedelta64(1, 's'),
          np.timedelta64('NaT', 's'),
          np.timedelta64(3, 's')], 2),
        ([np.timedelta64('NaT', 's')] * 3, 0),

        ([timedelta(days=5, seconds=14), timedelta(days=2, seconds=35),
          timedelta(days=-1, seconds=23)], 2),
        ([timedelta(days=1, seconds=43), timedelta(days=10, seconds=5),
          timedelta(days=5, seconds=14)], 0),
        ([timedelta(days=10, seconds=24), timedelta(days=10, seconds=5),
          timedelta(days=10, seconds=43)], 1),

        ([True, True, True, True, False], 4),
        ([True, True, True, False, True], 3),
        ([False, True, True, True, True], 0),
        ([False, True, False, True, True], 0),
    ]

    @pytest.mark.parametrize('data', nan_arr)
    def test_combinations(self, data):
        arr, pos = data
        with suppress_warnings() as sup:
            sup.filter(RuntimeWarning,
                       "invalid value encountered in reduce")
            min_val = np.min(arr)

        assert_equal(np.argmin(arr), pos, err_msg=f"{arr!r}")
        assert_equal(arr[np.argmin(arr)], min_val, err_msg=f"{arr!r}")

        # add padding to test SIMD loops
        rarr = np.repeat(arr, 129)
        rpos = pos * 129
        assert_equal(np.argmin(rarr), rpos, err_msg=f"{rarr!r}")
        assert_equal(rarr[np.argmin(rarr)], min_val, err_msg=f"{rarr!r}")

        padd = np.repeat(np.max(arr), 513)
        rarr = np.concatenate((arr, padd))
        rpos = pos
        assert_equal(np.argmin(rarr), rpos, err_msg=f"{rarr!r}")
        assert_equal(rarr[np.argmin(rarr)], min_val, err_msg=f"{rarr!r}")

    def test_minimum_signed_integers(self):

        a = np.array([1, -2**7, -2**7 + 1, 2**7 - 1], dtype=np.int8)
        assert_equal(np.argmin(a), 1)
        a = a.repeat(129)
        assert_equal(np.argmin(a), 129)

        a = np.array([1, -2**15, -2**15 + 1, 2**15 - 1], dtype=np.int16)
        assert_equal(np.argmin(a), 1)
        a = a.repeat(129)
        assert_equal(np.argmin(a), 129)

        a = np.array([1, -2**31, -2**31 + 1, 2**31 - 1], dtype=np.int32)
        assert_equal(np.argmin(a), 1)
        a = a.repeat(129)
        assert_equal(np.argmin(a), 129)

        a = np.array([1, -2**63, -2**63 + 1, 2**63 - 1], dtype=np.int64)
        assert_equal(np.argmin(a), 1)
        a = a.repeat(129)
        assert_equal(np.argmin(a), 129)

class TestMinMax:

    def test_scalar(self):
        assert_raises(AxisError, np.amax, 1, 1)
        assert_raises(AxisError, np.amin, 1, 1)

        assert_equal(np.amax(1, axis=0), 1)
        assert_equal(np.amin(1, axis=0), 1)
        assert_equal(np.amax(1, axis=None), 1)
        assert_equal(np.amin(1, axis=None), 1)

    def test_axis(self):
        assert_raises(AxisError, np.amax, [1, 2, 3], 1000)
        assert_equal(np.amax([[1, 2, 3]], axis=1), 3)

    def test_datetime(self):
        # Do not ignore NaT
        for dtype in ('m8[s]', 'm8[Y]'):
            a = np.arange(10).astype(dtype)
            assert_equal(np.amin(a), a[0])
            assert_equal(np.amax(a), a[9])
            a[3] = 'NaT'
            assert_equal(np.amin(a), a[3])
            assert_equal(np.amax(a), a[3])


class TestNewaxis:
    def test_basic(self):
        sk = np.array([0, -0.1, 0.1])
        res = 250 * sk[:, np.newaxis]
        assert_almost_equal(res.ravel(), 250 * sk)


class TestClip:
    def _check_range(self, x, cmin, cmax):
        assert_(np.all(x >= cmin))
        assert_(np.all(x <= cmax))

    def _clip_type(self, type_group, array_max,
                   clip_min, clip_max, inplace=False,
                   expected_min=None, expected_max=None):
        if expected_min is None:
            expected_min = clip_min
        if expected_max is None:
            expected_max = clip_max

        for T in np._core.sctypes[type_group]:
            if sys.byteorder == 'little':
                byte_orders = ['=', '>']
            else:
                byte_orders = ['<', '=']

            for byteorder in byte_orders:
                dtype = np.dtype(T).newbyteorder(byteorder)

                x = (np.random.random(1000) * array_max).astype(dtype)
                if inplace:
                    # The tests that call us pass clip_min and clip_max that
                    # might not fit in the destination dtype. They were written
                    # assuming the previous unsafe casting, which now must be
                    # passed explicitly to avoid a warning.
                    x.clip(clip_min, clip_max, x, casting='unsafe')
                else:
                    x = x.clip(clip_min, clip_max)
                    byteorder = '='

                if x.dtype.byteorder == '|':
                    byteorder = '|'
                assert_equal(x.dtype.byteorder, byteorder)
                self._check_range(x, expected_min, expected_max)
        return x

    def test_basic(self):
        for inplace in [False, True]:
            self._clip_type(
                'float', 1024, -12.8, 100.2, inplace=inplace)
            self._clip_type(
                'float', 1024, 0, 0, inplace=inplace)

            self._clip_type(
                'int', 1024, -120, 100, inplace=inplace)
            self._clip_type(
                'int', 1024, 0, 0, inplace=inplace)

            self._clip_type(
                'uint', 1024, 0, 0, inplace=inplace)
            self._clip_type(
                'uint', 1024, 10, 100, inplace=inplace)

    @pytest.mark.parametrize("inplace", [False, True])
    def test_int_out_of_range(self, inplace):
        # Simple check for out-of-bound integers, also testing the in-place
        # path.
        x = (np.random.random(1000) * 255).astype("uint8")
        out = np.empty_like(x)
        res = x.clip(-1, 300, out=out if inplace else None)
        assert res is out or not inplace
        assert (res == x).all()

        res = x.clip(-1, 50, out=out if inplace else None)
        assert res is out or not inplace
        assert (res <= 50).all()
        assert (res[x <= 50] == x[x <= 50]).all()

        res = x.clip(100, 1000, out=out if inplace else None)
        assert res is out or not inplace
        assert (res >= 100).all()
        assert (res[x >= 100] == x[x >= 100]).all()

    def test_record_array(self):
        rec = np.array([(-5, 2.0, 3.0), (5.0, 4.0, 3.0)],
                       dtype=[('x', '<f8'), ('y', '<f8'), ('z', '<f8')])
        y = rec['x'].clip(-0.3, 0.5)
        self._check_range(y, -0.3, 0.5)

    def test_max_or_min(self):
        val = np.array([0, 1, 2, 3, 4, 5, 6, 7])
        x = val.clip(3)
        assert_(np.all(x >= 3))
        x = val.clip(min=3)
        assert_(np.all(x >= 3))
        x = val.clip(max=4)
        assert_(np.all(x <= 4))

    def test_nan(self):
        input_arr = np.array([-2., np.nan, 0.5, 3., 0.25, np.nan])
        result = input_arr.clip(-1, 1)
        expected = np.array([-1., np.nan, 0.5, 1., 0.25, np.nan])
        assert_array_equal(result, expected)


class TestCompress:
    def test_axis(self):
        tgt = [[5, 6, 7, 8, 9]]
        arr = np.arange(10).reshape(2, 5)
        out = np.compress([0, 1], arr, axis=0)
        assert_equal(out, tgt)

        tgt = [[1, 3], [6, 8]]
        out = np.compress([0, 1, 0, 1, 0], arr, axis=1)
        assert_equal(out, tgt)

    def test_truncate(self):
        tgt = [[1], [6]]
        arr = np.arange(10).reshape(2, 5)
        out = np.compress([0, 1], arr, axis=1)
        assert_equal(out, tgt)

    def test_flatten(self):
        arr = np.arange(10).reshape(2, 5)
        out = np.compress([0, 1], arr)
        assert_equal(out, 1)


class TestPutmask:
    def tst_basic(self, x, T, mask, val):
        np.putmask(x, mask, val)
        assert_equal(x[mask], np.array(val, T))

    def test_ip_types(self):
        unchecked_types = [bytes, str, np.void]

        x = np.random.random(1000) * 100
        mask = x < 40

        for val in [-100, 0, 15]:
            for types in np._core.sctypes.values():
                for T in types:
                    if T not in unchecked_types:
                        if val < 0 and np.dtype(T).kind == "u":
                            val = np.iinfo(T).max - 99
                        self.tst_basic(x.copy().astype(T), T, mask, val)

            # Also test string of a length which uses an untypical length
            dt = np.dtype("S3")
            self.tst_basic(x.astype(dt), dt.type, mask, dt.type(val)[:3])

    def test_mask_size(self):
        assert_raises(ValueError, np.putmask, np.array([1, 2, 3]), [True], 5)

    @pytest.mark.parametrize('dtype', ('>i4', '<i4'))
    def test_byteorder(self, dtype):
        x = np.array([1, 2, 3], dtype)
        np.putmask(x, [True, False, True], -1)
        assert_array_equal(x, [-1, 2, -1])

    def test_record_array(self):
        # Note mixed byteorder.
        rec = np.array([(-5, 2.0, 3.0), (5.0, 4.0, 3.0)],
                      dtype=[('x', '<f8'), ('y', '>f8'), ('z', '<f8')])
        np.putmask(rec['x'], [True, False], 10)
        assert_array_equal(rec['x'], [10, 5])
        assert_array_equal(rec['y'], [2, 4])
        assert_array_equal(rec['z'], [3, 3])
        np.putmask(rec['y'], [True, False], 11)
        assert_array_equal(rec['x'], [10, 5])
        assert_array_equal(rec['y'], [11, 4])
        assert_array_equal(rec['z'], [3, 3])

    def test_overlaps(self):
        # gh-6272 check overlap
        x = np.array([True, False, True, False])
        np.putmask(x[1:4], [True, True, True], x[:3])
        assert_equal(x, np.array([True, True, False, True]))

        x = np.array([True, False, True, False])
        np.putmask(x[1:4], x[:3], [True, False, True])
        assert_equal(x, np.array([True, True, True, True]))

    def test_writeable(self):
        a = np.arange(5)
        a.flags.writeable = False

        with pytest.raises(ValueError):
            np.putmask(a, a >= 2, 3)

    def test_kwargs(self):
        x = np.array([0, 0])
        np.putmask(x, [0, 1], [-1, -2])
        assert_array_equal(x, [0, -2])

        x = np.array([0, 0])
        np.putmask(x, mask=[0, 1], values=[-1, -2])
        assert_array_equal(x, [0, -2])

        x = np.array([0, 0])
        np.putmask(x, values=[-1, -2],  mask=[0, 1])
        assert_array_equal(x, [0, -2])

        with pytest.raises(TypeError):
            np.putmask(a=x, values=[-1, -2],  mask=[0, 1])


class TestTake:
    def tst_basic(self, x):
        ind = list(range(x.shape[0]))
        assert_array_equal(x.take(ind, axis=0), x)

    def test_ip_types(self):
        unchecked_types = [bytes, str, np.void]

        x = np.random.random(24) * 100
        x.shape = 2, 3, 4
        for types in np._core.sctypes.values():
            for T in types:
                if T not in unchecked_types:
                    self.tst_basic(x.copy().astype(T))

            # Also test string of a length which uses an untypical length
            self.tst_basic(x.astype("S3"))

    def test_raise(self):
        x = np.random.random(24) * 100
        x.shape = 2, 3, 4
        assert_raises(IndexError, x.take, [0, 1, 2], axis=0)
        assert_raises(IndexError, x.take, [-3], axis=0)
        assert_array_equal(x.take([-1], axis=0)[0], x[1])

    def test_clip(self):
        x = np.random.random(24) * 100
        x.shape = 2, 3, 4
        assert_array_equal(x.take([-1], axis=0, mode='clip')[0], x[0])
        assert_array_equal(x.take([2], axis=0, mode='clip')[0], x[1])

    def test_wrap(self):
        x = np.random.random(24) * 100
        x.shape = 2, 3, 4
        assert_array_equal(x.take([-1], axis=0, mode='wrap')[0], x[1])
        assert_array_equal(x.take([2], axis=0, mode='wrap')[0], x[0])
        assert_array_equal(x.take([3], axis=0, mode='wrap')[0], x[1])

    @pytest.mark.parametrize('dtype', ('>i4', '<i4'))
    def test_byteorder(self, dtype):
        x = np.array([1, 2, 3], dtype)
        assert_array_equal(x.take([0, 2, 1]), [1, 3, 2])

    def test_record_array(self):
        # Note mixed byteorder.
        rec = np.array([(-5, 2.0, 3.0), (5.0, 4.0, 3.0)],
                      dtype=[('x', '<f8'), ('y', '>f8'), ('z', '<f8')])
        rec1 = rec.take([1])
        assert_(rec1['x'] == 5.0 and rec1['y'] == 4.0)

    def test_out_overlap(self):
        # gh-6272 check overlap on out
        x = np.arange(5)
        y = np.take(x, [1, 2, 3], out=x[2:5], mode='wrap')
        assert_equal(y, np.array([1, 2, 3]))

    @pytest.mark.parametrize('shape', [(1, 2), (1,), ()])
    def test_ret_is_out(self, shape):
        # 0d arrays should not be an exception to this rule
        x = np.arange(5)
        inds = np.zeros(shape, dtype=np.intp)
        out = np.zeros(shape, dtype=x.dtype)
        ret = np.take(x, inds, out=out)
        assert ret is out


class TestLexsort:
    @pytest.mark.parametrize('dtype', [
        np.uint8, np.uint16, np.uint32, np.uint64,
        np.int8, np.int16, np.int32, np.int64,
        np.float16, np.float32, np.float64
    ])
    def test_basic(self, dtype):
        a = np.array([1, 2, 1, 3, 1, 5], dtype=dtype)
        b = np.array([0, 4, 5, 6, 2, 3], dtype=dtype)
        idx = np.lexsort((b, a))
        expected_idx = np.array([0, 4, 2, 1, 3, 5])
        assert_array_equal(idx, expected_idx)
        assert_array_equal(a[idx], np.sort(a))

    def test_mixed(self):
        a = np.array([1, 2, 1, 3, 1, 5])
        b = np.array([0, 4, 5, 6, 2, 3], dtype='datetime64[D]')

        idx = np.lexsort((b, a))
        expected_idx = np.array([0, 4, 2, 1, 3, 5])
        assert_array_equal(idx, expected_idx)

    def test_datetime(self):
        a = np.array([0, 0, 0], dtype='datetime64[D]')
        b = np.array([2, 1, 0], dtype='datetime64[D]')
        idx = np.lexsort((b, a))
        expected_idx = np.array([2, 1, 0])
        assert_array_equal(idx, expected_idx)

        a = np.array([0, 0, 0], dtype='timedelta64[D]')
        b = np.array([2, 1, 0], dtype='timedelta64[D]')
        idx = np.lexsort((b, a))
        expected_idx = np.array([2, 1, 0])
        assert_array_equal(idx, expected_idx)

    def test_object(self):  # gh-6312
        a = np.random.choice(10, 1000)
        b = np.random.choice(['abc', 'xy', 'wz', 'efghi', 'qwst', 'x'], 1000)

        for u in a, b:
            left = np.lexsort((u.astype('O'),))
            right = np.argsort(u, kind='mergesort')
            assert_array_equal(left, right)

        for u, v in (a, b), (b, a):
            idx = np.lexsort((u, v))
            assert_array_equal(idx, np.lexsort((u.astype('O'), v)))
            assert_array_equal(idx, np.lexsort((u, v.astype('O'))))
            u, v = np.array(u, dtype='object'), np.array(v, dtype='object')
            assert_array_equal(idx, np.lexsort((u, v)))

    def test_strings(self):  # gh-27984
        for dtype in "TU":
            surnames = np.array(['Hertz',    'Galilei', 'Hertz'], dtype=dtype)
            first_names = np.array(['Heinrich', 'Galileo', 'Gustav'], dtype=dtype)
            assert_array_equal(np.lexsort((first_names, surnames)), [1, 2, 0])

    def test_invalid_axis(self):  # gh-7528
        x = np.linspace(0., 1., 42 * 3).reshape(42, 3)
        assert_raises(AxisError, np.lexsort, x, axis=2)

class TestIO:
    """Test tofile, fromfile, tobytes, and fromstring"""

    @pytest.fixture()
    def x(self):
        shape = (2, 4, 3)
        rand = np.random.random
        x = rand(shape) + rand(shape).astype(complex) * 1j
        x[0, :, 1] = [np.nan, np.inf, -np.inf, np.nan]
        return x

    @pytest.fixture(params=["string", "path_obj"])
    def tmp_filename(self, tmp_path, request):
        # This fixture covers two cases:
        # one where the filename is a string and
        # another where it is a pathlib object
        filename = tmp_path / "file"
        if request.param == "string":
            filename = str(filename)
        yield filename

    def test_nofile(self):
        # this should probably be supported as a file
        # but for now test for proper errors
        b = io.BytesIO()
        assert_raises(OSError, np.fromfile, b, np.uint8, 80)
        d = np.ones(7)
        assert_raises(OSError, lambda x: x.tofile(b), d)

    def test_bool_fromstring(self):
        v = np.array([True, False, True, False], dtype=np.bool)
        y = np.fromstring('1 0 -2.3 0.0', sep=' ', dtype=np.bool)
        assert_array_equal(v, y)

    def test_uint64_fromstring(self):
        d = np.fromstring("9923372036854775807 104783749223640",
                          dtype=np.uint64, sep=' ')
        e = np.array([9923372036854775807, 104783749223640], dtype=np.uint64)
        assert_array_equal(d, e)

    def test_int64_fromstring(self):
        d = np.fromstring("-25041670086757 104783749223640",
                          dtype=np.int64, sep=' ')
        e = np.array([-25041670086757, 104783749223640], dtype=np.int64)
        assert_array_equal(d, e)

    def test_fromstring_count0(self):
        d = np.fromstring("1,2", sep=",", dtype=np.int64, count=0)
        assert d.shape == (0,)

    def test_empty_files_text(self, tmp_filename):
        with open(tmp_filename, 'w') as f:
            pass
        y = np.fromfile(tmp_filename)
        assert_(y.size == 0, "Array not empty")

    def test_empty_files_binary(self, tmp_filename):
        with open(tmp_filename, 'wb') as f:
            pass
        y = np.fromfile(tmp_filename, sep=" ")
        assert_(y.size == 0, "Array not empty")

    def test_roundtrip_file(self, x, tmp_filename):
        with open(tmp_filename, 'wb') as f:
            x.tofile(f)
        # NB. doesn't work with flush+seek, due to use of C stdio
        with open(tmp_filename, 'rb') as f:
            y = np.fromfile(f, dtype=x.dtype)
        assert_array_equal(y, x.flat)

    def test_roundtrip(self, x, tmp_filename):
        x.tofile(tmp_filename)
        y = np.fromfile(tmp_filename, dtype=x.dtype)
        assert_array_equal(y, x.flat)

    def test_roundtrip_dump_pathlib(self, x, tmp_filename):
        p = pathlib.Path(tmp_filename)
        x.dump(p)
        y = np.load(p, allow_pickle=True)
        assert_array_equal(y, x)

    def test_roundtrip_binary_str(self, x):
        s = x.tobytes()
        y = np.frombuffer(s, dtype=x.dtype)
        assert_array_equal(y, x.flat)

        s = x.tobytes('F')
        y = np.frombuffer(s, dtype=x.dtype)
        assert_array_equal(y, x.flatten('F'))

    def test_roundtrip_str(self, x):
        x = x.real.ravel()
        s = "@".join(map(str, x))
        y = np.fromstring(s, sep="@")
        nan_mask = ~np.isfinite(x)
        assert_array_equal(x[nan_mask], y[nan_mask])
        assert_array_equal(x[~nan_mask], y[~nan_mask])

    def test_roundtrip_repr(self, x):
        x = x.real.ravel()
        s = "@".join(repr(x)[11:-1] for x in x)
        y = np.fromstring(s, sep="@")
        assert_array_equal(x, y)

    def test_unseekable_fromfile(self, x, tmp_filename):
        # gh-6246
        x.tofile(tmp_filename)

        def fail(*args, **kwargs):
            raise OSError('Can not tell or seek')

        with open(tmp_filename, 'rb', buffering=0) as f:
            f.seek = fail
            f.tell = fail
            assert_raises(OSError, np.fromfile, f, dtype=x.dtype)

    def test_io_open_unbuffered_fromfile(self, x, tmp_filename):
        # gh-6632
        x.tofile(tmp_filename)
        with open(tmp_filename, 'rb', buffering=0) as f:
            y = np.fromfile(f, dtype=x.dtype)
            assert_array_equal(y, x.flat)

    def test_largish_file(self, tmp_filename):
        # check the fallocate path on files > 16MB
        d = np.zeros(4 * 1024 ** 2)
        d.tofile(tmp_filename)
        assert_equal(os.path.getsize(tmp_filename), d.nbytes)
        assert_array_equal(d, np.fromfile(tmp_filename))
        # check offset
        with open(tmp_filename, "r+b") as f:
            f.seek(d.nbytes)
            d.tofile(f)
            assert_equal(os.path.getsize(tmp_filename), d.nbytes * 2)
        # check append mode (gh-8329)
        open(tmp_filename, "w").close()  # delete file contents
        with open(tmp_filename, "ab") as f:
            d.tofile(f)
        assert_array_equal(d, np.fromfile(tmp_filename))
        with open(tmp_filename, "ab") as f:
            d.tofile(f)
        assert_equal(os.path.getsize(tmp_filename), d.nbytes * 2)

    def test_io_open_buffered_fromfile(self, x, tmp_filename):
        # gh-6632
        x.tofile(tmp_filename)
        with open(tmp_filename, 'rb', buffering=-1) as f:
            y = np.fromfile(f, dtype=x.dtype)
        assert_array_equal(y, x.flat)

    def test_file_position_after_fromfile(self, tmp_filename):
        # gh-4118
        sizes = [io.DEFAULT_BUFFER_SIZE // 8,
                 io.DEFAULT_BUFFER_SIZE,
                 io.DEFAULT_BUFFER_SIZE * 8]

        for size in sizes:
            with open(tmp_filename, 'wb') as f:
                f.seek(size - 1)
                f.write(b'\0')

            for mode in ['rb', 'r+b']:
                err_msg = "%d %s" % (size, mode)

                with open(tmp_filename, mode) as f:
                    f.read(2)
                    np.fromfile(f, dtype=np.float64, count=1)
                    pos = f.tell()
                assert_equal(pos, 10, err_msg=err_msg)

    def test_file_position_after_tofile(self, tmp_filename):
        # gh-4118
        sizes = [io.DEFAULT_BUFFER_SIZE // 8,
                 io.DEFAULT_BUFFER_SIZE,
                 io.DEFAULT_BUFFER_SIZE * 8]

        for size in sizes:
            err_msg = "%d" % (size,)

            with open(tmp_filename, 'wb') as f:
                f.seek(size - 1)
                f.write(b'\0')
                f.seek(10)
                f.write(b'12')
                np.array([0], dtype=np.float64).tofile(f)
                pos = f.tell()
            assert_equal(pos, 10 + 2 + 8, err_msg=err_msg)

            with open(tmp_filename, 'r+b') as f:
                f.read(2)
                f.seek(0, 1)  # seek between read&write required by ANSI C
                np.array([0], dtype=np.float64).tofile(f)
                pos = f.tell()
            assert_equal(pos, 10, err_msg=err_msg)

    def test_load_object_array_fromfile(self, tmp_filename):
        # gh-12300
        with open(tmp_filename, 'w') as f:
            # Ensure we have a file with consistent contents
            pass

        with open(tmp_filename, 'rb') as f:
            assert_raises_regex(ValueError, "Cannot read into object array",
                                np.fromfile, f, dtype=object)

        assert_raises_regex(ValueError, "Cannot read into object array",
                            np.fromfile, tmp_filename, dtype=object)

    def test_fromfile_offset(self, x, tmp_filename):
        with open(tmp_filename, 'wb') as f:
            x.tofile(f)

        with open(tmp_filename, 'rb') as f:
            y = np.fromfile(f, dtype=x.dtype, offset=0)
            assert_array_equal(y, x.flat)

        with open(tmp_filename, 'rb') as f:
            count_items = len(x.flat) // 8
            offset_items = len(x.flat) // 4
            offset_bytes = x.dtype.itemsize * offset_items
            y = np.fromfile(
                f, dtype=x.dtype, count=count_items, offset=offset_bytes
            )
            assert_array_equal(
                y, x.flat[offset_items:offset_items + count_items]
            )

            # subsequent seeks should stack
            offset_bytes = x.dtype.itemsize
            z = np.fromfile(f, dtype=x.dtype, offset=offset_bytes)
            assert_array_equal(z, x.flat[offset_items + count_items + 1:])

        with open(tmp_filename, 'wb') as f:
            x.tofile(f, sep=",")

        with open(tmp_filename, 'rb') as f:
            assert_raises_regex(
                    TypeError,
                    "'offset' argument only permitted for binary files",
                    np.fromfile, tmp_filename, dtype=x.dtype,
                    sep=",", offset=1)

    @pytest.mark.skipif(IS_PYPY, reason="bug in PyPy's PyNumber_AsSsize_t")
    def test_fromfile_bad_dup(self, x, tmp_filename):
        def dup_str(fd):
            return 'abc'

        def dup_bigint(fd):
            return 2**68

        old_dup = os.dup
        try:
            with open(tmp_filename, 'wb') as f:
                x.tofile(f)
                for dup, exc in ((dup_str, TypeError), (dup_bigint, OSError)):
                    os.dup = dup
                    assert_raises(exc, np.fromfile, f)
        finally:
            os.dup = old_dup

    def _check_from(self, s, value, filename, **kw):
        if 'sep' not in kw:
            y = np.frombuffer(s, **kw)
        else:
            y = np.fromstring(s, **kw)
        assert_array_equal(y, value)

        with open(filename, 'wb') as f:
            f.write(s)
        y = np.fromfile(filename, **kw)
        assert_array_equal(y, value)

    @pytest.fixture(params=["period", "comma"])
    def decimal_sep_localization(self, request):
        """
        Including this fixture in a test will automatically
        execute it with both types of decimal separator.

        So::

            def test_decimal(decimal_sep_localization):
                pass

        is equivalent to the following two tests::

            def test_decimal_period_separator():
                pass

            def test_decimal_comma_separator():
                with CommaDecimalPointLocale():
                    pass
        """
        if request.param == "period":
            yield
        elif request.param == "comma":
            with CommaDecimalPointLocale():
                yield
        else:
            assert False, request.param

    def test_nan(self, tmp_filename, decimal_sep_localization):
        self._check_from(
            b"nan +nan -nan NaN nan(foo) +NaN(BAR) -NAN(q_u_u_x_)",
            [np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan],
            tmp_filename,
            sep=' ')

    def test_inf(self, tmp_filename, decimal_sep_localization):
        self._check_from(
            b"inf +inf -inf infinity -Infinity iNfInItY -inF",
            [np.inf, np.inf, -np.inf, np.inf, -np.inf, np.inf, -np.inf],
            tmp_filename,
            sep=' ')

    def test_numbers(self, tmp_filename, decimal_sep_localization):
        self._check_from(
            b"1.234 -1.234 .3 .3e55 -123133.1231e+133",
            [1.234, -1.234, .3, .3e55, -123133.1231e+133],
            tmp_filename,
            sep=' ')

    def test_binary(self, tmp_filename):
        self._check_from(
            b'\x00\x00\x80?\x00\x00\x00@\x00\x00@@\x00\x00\x80@',
            np.array([1, 2, 3, 4]),
            tmp_filename,
            dtype='<f4')

    def test_string(self, tmp_filename):
        self._check_from(b'1,2,3,4', [1., 2., 3., 4.], tmp_filename, sep=',')

    def test_counted_string(self, tmp_filename, decimal_sep_localization):
        self._check_from(
            b'1,2,3,4', [1., 2., 3., 4.], tmp_filename, count=4, sep=',')
        self._check_from(
            b'1,2,3,4', [1., 2., 3.], tmp_filename, count=3, sep=',')
        self._check_from(
            b'1,2,3,4', [1., 2., 3., 4.], tmp_filename, count=-1, sep=',')

    def test_string_with_ws(self, tmp_filename):
        self._check_from(
            b'1 2  3     4   ', [1, 2, 3, 4], tmp_filename, dtype=int, sep=' ')

    def test_counted_string_with_ws(self, tmp_filename):
        self._check_from(
            b'1 2  3     4   ', [1, 2, 3], tmp_filename, count=3, dtype=int,
            sep=' ')

    def test_ascii(self, tmp_filename, decimal_sep_localization):
        self._check_from(
            b'1 , 2 , 3 , 4', [1., 2., 3., 4.], tmp_filename, sep=',')
        self._check_from(
            b'1,2,3,4', [1., 2., 3., 4.], tmp_filename, dtype=float, sep=',')

    def test_malformed(self, tmp_filename, decimal_sep_localization):
        with assert_raises(ValueError):
            self._check_from(
                b'1.234 1,234', [1.234, 1.], tmp_filename, sep=' ')

    def test_long_sep(self, tmp_filename):
        self._check_from(
            b'1_x_3_x_4_x_5', [1, 3, 4, 5], tmp_filename, sep='_x_')

    def test_dtype(self, tmp_filename):
        v = np.array([1, 2, 3, 4], dtype=np.int_)
        self._check_from(b'1,2,3,4', v, tmp_filename, sep=',', dtype=np.int_)

    def test_dtype_bool(self, tmp_filename):
        # can't use _check_from because fromstring can't handle True/False
        v = np.array([True, False, True, False], dtype=np.bool)
        s = b'1,0,-2.3,0'
        with open(tmp_filename, 'wb') as f:
            f.write(s)
        y = np.fromfile(tmp_filename, sep=',', dtype=np.bool)
        assert_(y.dtype == '?')
        assert_array_equal(y, v)

    def test_tofile_sep(self, tmp_filename, decimal_sep_localization):
        x = np.array([1.51, 2, 3.51, 4], dtype=float)
        with open(tmp_filename, 'w') as f:
            x.tofile(f, sep=',')
        with open(tmp_filename, 'r') as f:
            s = f.read()
        #assert_equal(s, '1.51,2.0,3.51,4.0')
        y = np.array([float(p) for p in s.split(',')])
        assert_array_equal(x, y)

    def test_tofile_format(self, tmp_filename, decimal_sep_localization):
        x = np.array([1.51, 2, 3.51, 4], dtype=float)
        with open(tmp_filename, 'w') as f:
            x.tofile(f, sep=',', format='%.2f')
        with open(tmp_filename, 'r') as f:
            s = f.read()
        assert_equal(s, '1.51,2.00,3.51,4.00')

    def test_tofile_cleanup(self, tmp_filename):
        x = np.zeros((10), dtype=object)
        with open(tmp_filename, 'wb') as f:
            assert_raises(OSError, lambda: x.tofile(f, sep=''))
        # Dup-ed file handle should be closed or remove will fail on Windows OS
        os.remove(tmp_filename)

        # Also make sure that we close the Python handle
        assert_raises(OSError, lambda: x.tofile(tmp_filename))
        os.remove(tmp_filename)

    def test_fromfile_subarray_binary(self, tmp_filename):
        # Test subarray dtypes which are absorbed into the shape
        x = np.arange(24, dtype="i4").reshape(2, 3, 4)
        x.tofile(tmp_filename)
        res = np.fromfile(tmp_filename, dtype="(3,4)i4")
        assert_array_equal(x, res)

        x_str = x.tobytes()
        with pytest.raises(ValueError):
            # binary fromstring raises
            np.fromstring(x_str, dtype="(3,4)i4")

    def test_parsing_subarray_unsupported(self, tmp_filename):
        # We currently do not support parsing subarray dtypes
        data = "12,42,13," * 50
        with pytest.raises(ValueError):
            expected = np.fromstring(data, dtype="(3,)i", sep=",")

        with open(tmp_filename, "w") as f:
            f.write(data)

        with pytest.raises(ValueError):
            np.fromfile(tmp_filename, dtype="(3,)i", sep=",")

    def test_read_shorter_than_count_subarray(self, tmp_filename):
        # Test that requesting more values does not cause any problems
        # in conjunction with subarray dimensions being absorbed into the
        # array dimension.
        expected = np.arange(511 * 10, dtype="i").reshape(-1, 10)

        binary = expected.tobytes()
        with pytest.raises(ValueError):
            np.fromstring(binary, dtype="(10,)i", count=10000)

        expected.tofile(tmp_filename)
        res = np.fromfile(tmp_filename, dtype="(10,)i", count=10000)
        assert_array_equal(res, expected)


class TestFromBuffer:
    @pytest.mark.parametrize('byteorder', ['<', '>'])
    @pytest.mark.parametrize('dtype', [float, int, complex])
    def test_basic(self, byteorder, dtype):
        dt = np.dtype(dtype).newbyteorder(byteorder)
        x = (np.random.random((4, 7)) * 5).astype(dt)
        buf = x.tobytes()
        assert_array_equal(np.frombuffer(buf, dtype=dt), x.flat)

    @pytest.mark.parametrize("obj", [np.arange(10), b"12345678"])
    def test_array_base(self, obj):
        # Objects (including NumPy arrays), which do not use the
        # `release_buffer` slot should be directly used as a base object.
        # See also gh-21612
        new = np.frombuffer(obj)
        assert new.base is obj

    def test_empty(self):
        assert_array_equal(np.frombuffer(b''), np.array([]))

    @pytest.mark.skipif(IS_PYPY,
            reason="PyPy's memoryview currently does not track exports. See: "
                   "https://foss.heptapod.net/pypy/pypy/-/issues/3724")
    def test_mmap_close(self):
        # The old buffer protocol was not safe for some things that the new
        # one is.  But `frombuffer` always used the old one for a long time.
        # Checks that it is safe with the new one (using memoryviews)
        with tempfile.TemporaryFile(mode='wb') as tmp:
            tmp.write(b"asdf")
            tmp.flush()
            mm = mmap.mmap(tmp.fileno(), 0)
            arr = np.frombuffer(mm, dtype=np.uint8)
            with pytest.raises(BufferError):
                mm.close()  # cannot close while array uses the buffer
            del arr
            mm.close()

class TestFlat:
    def setup_method(self):
        a0 = np.arange(20.0)
        a = a0.reshape(4, 5)
        a0.shape = (4, 5)
        a.flags.writeable = False
        self.a = a
        self.b = a[::2, ::2]
        self.a0 = a0
        self.b0 = a0[::2, ::2]

    def test_contiguous(self):
        testpassed = False
        try:
            self.a.flat[12] = 100.0
        except ValueError:
            testpassed = True
        assert_(testpassed)
        assert_(self.a.flat[12] == 12.0)

    def test_discontiguous(self):
        testpassed = False
        try:
            self.b.flat[4] = 100.0
        except ValueError:
            testpassed = True
        assert_(testpassed)
        assert_(self.b.flat[4] == 12.0)

    def test___array__(self):
        c = self.a.flat.__array__()
        d = self.b.flat.__array__()
        e = self.a0.flat.__array__()
        f = self.b0.flat.__array__()

        assert_(c.flags.writeable is False)
        assert_(d.flags.writeable is False)
        assert_(e.flags.writeable is True)
        assert_(f.flags.writeable is False)
        assert_(c.flags.writebackifcopy is False)
        assert_(d.flags.writebackifcopy is False)
        assert_(e.flags.writebackifcopy is False)
        assert_(f.flags.writebackifcopy is False)

    @pytest.mark.skipif(not HAS_REFCOUNT, reason="Python lacks refcounts")
    def test_refcount(self):
        # includes regression test for reference count error gh-13165
        inds = [np.intp(0), np.array([True] * self.a.size), np.array([0]), None]
        indtype = np.dtype(np.intp)
        rc_indtype = sys.getrefcount(indtype)
        for ind in inds:
            rc_ind = sys.getrefcount(ind)
            for _ in range(100):
                try:
                    self.a.flat[ind]
                except IndexError:
                    pass
            assert_(abs(sys.getrefcount(ind) - rc_ind) < 50)
            assert_(abs(sys.getrefcount(indtype) - rc_indtype) < 50)

    def test_index_getset(self):
        it = np.arange(10).reshape(2, 1, 5).flat
        with pytest.raises(AttributeError):
            it.index = 10

        for _ in it:
            pass
        # Check the value of `.index` is updated correctly (see also gh-19153)
        # If the type was incorrect, this would show up on big-endian machines
        assert it.index == it.base.size

    def test_maxdims(self):
        # The flat iterator and thus attribute is currently unfortunately
        # limited to only 32 dimensions (after bumping it to 64 for 2.0)
        a = np.ones((1,) * 64)

        with pytest.raises(RuntimeError,
                match=".*32 dimensions but the array has 64"):
            a.flat


class TestResize:

    @_no_tracing
    def test_basic(self):
        x = np.array([[1, 0, 0], [0, 1, 0], [0, 0, 1]])
        if IS_PYPY:
            x.resize((5, 5), refcheck=False)
        else:
            x.resize((5, 5))
        assert_array_equal(x.flat[:9],
                np.array([[1, 0, 0], [0, 1, 0], [0, 0, 1]]).flat)
        assert_array_equal(x[9:].flat, 0)

    def test_check_reference(self):
        x = np.array([[1, 0, 0], [0, 1, 0], [0, 0, 1]])
        y = x
        assert_raises(ValueError, x.resize, (5, 1))

    @_no_tracing
    def test_int_shape(self):
        x = np.eye(3)
        if IS_PYPY:
            x.resize(3, refcheck=False)
        else:
            x.resize(3)
        assert_array_equal(x, np.eye(3)[0, :])

    def test_none_shape(self):
        x = np.eye(3)
        x.resize(None)
        assert_array_equal(x, np.eye(3))
        x.resize()
        assert_array_equal(x, np.eye(3))

    def test_0d_shape(self):
        # to it multiple times to test it does not break alloc cache gh-9216
        for i in range(10):
            x = np.empty((1,))
            x.resize(())
            assert_equal(x.shape, ())
            assert_equal(x.size, 1)
            x = np.empty(())
            x.resize((1,))
            assert_equal(x.shape, (1,))
            assert_equal(x.size, 1)

    def test_invalid_arguments(self):
        assert_raises(TypeError, np.eye(3).resize, 'hi')
        assert_raises(ValueError, np.eye(3).resize, -1)
        assert_raises(TypeError, np.eye(3).resize, order=1)
        assert_raises(TypeError, np.eye(3).resize, refcheck='hi')

    @_no_tracing
    def test_freeform_shape(self):
        x = np.eye(3)
        if IS_PYPY:
            x.resize(3, 2, 1, refcheck=False)
        else:
            x.resize(3, 2, 1)
        assert_(x.shape == (3, 2, 1))

    @_no_tracing
    def test_zeros_appended(self):
        x = np.eye(3)
        if IS_PYPY:
            x.resize(2, 3, 3, refcheck=False)
        else:
            x.resize(2, 3, 3)
        assert_array_equal(x[0], np.eye(3))
        assert_array_equal(x[1], np.zeros((3, 3)))

    @_no_tracing
    def test_obj_obj(self):
        # check memory is initialized on resize, gh-4857
        a = np.ones(10, dtype=[('k', object, 2)])
        if IS_PYPY:
            a.resize(15, refcheck=False)
        else:
            a.resize(15,)
        assert_equal(a.shape, (15,))
        assert_array_equal(a['k'][-5:], 0)
        assert_array_equal(a['k'][:-5], 1)

    def test_empty_view(self):
        # check that sizes containing a zero don't trigger a reallocate for
        # already empty arrays
        x = np.zeros((10, 0), int)
        x_view = x[...]
        x_view.resize((0, 10))
        x_view.resize((0, 100))

    def test_check_weakref(self):
        x = np.array([[1, 0, 0], [0, 1, 0], [0, 0, 1]])
        xref = weakref.ref(x)
        assert_raises(ValueError, x.resize, (5, 1))


class TestRecord:
    def test_field_rename(self):
        dt = np.dtype([('f', float), ('i', int)])
        dt.names = ['p', 'q']
        assert_equal(dt.names, ['p', 'q'])

    def test_multiple_field_name_occurrence(self):
        def test_dtype_init():
            np.dtype([("A", "f8"), ("B", "f8"), ("A", "f8")])

        # Error raised when multiple fields have the same name
        assert_raises(ValueError, test_dtype_init)

    def test_bytes_fields(self):
        # Bytes are not allowed in field names and not recognized in titles
        # on Py3
        assert_raises(TypeError, np.dtype, [(b'a', int)])
        assert_raises(TypeError, np.dtype, [(('b', b'a'), int)])

        dt = np.dtype([((b'a', 'b'), int)])
        assert_raises(TypeError, dt.__getitem__, b'a')

        x = np.array([(1,), (2,), (3,)], dtype=dt)
        assert_raises(IndexError, x.__getitem__, b'a')

        y = x[0]
        assert_raises(IndexError, y.__getitem__, b'a')

    def test_multiple_field_name_unicode(self):
        def test_dtype_unicode():
            np.dtype([("\u20B9", "f8"), ("B", "f8"), ("\u20B9", "f8")])

        # Error raised when multiple fields have the same name(unicode included)
        assert_raises(ValueError, test_dtype_unicode)

    def test_fromarrays_unicode(self):
        # A single name string provided to fromarrays() is allowed to be unicode
        x = np._core.records.fromarrays(
            [[0], [1]], names='a,b', formats='i4,i4')
        assert_equal(x['a'][0], 0)
        assert_equal(x['b'][0], 1)

    def test_unicode_order(self):
        # Test that we can sort with order as a unicode field name
        name = 'b'
        x = np.array([1, 3, 2], dtype=[(name, int)])
        x.sort(order=name)
        assert_equal(x['b'], np.array([1, 2, 3]))

    def test_field_names(self):
        # Test unicode and 8-bit / byte strings can be used
        a = np.zeros((1,), dtype=[('f1', 'i4'),
                                  ('f2', 'i4'),
                                  ('f3', [('sf1', 'i4')])])
        # byte string indexing fails gracefully
        assert_raises(IndexError, a.__setitem__, b'f1', 1)
        assert_raises(IndexError, a.__getitem__, b'f1')
        assert_raises(IndexError, a['f1'].__setitem__, b'sf1', 1)
        assert_raises(IndexError, a['f1'].__getitem__, b'sf1')
        b = a.copy()
        fn1 = 'f1'
        b[fn1] = 1
        assert_equal(b[fn1], 1)
        fnn = 'not at all'
        assert_raises(ValueError, b.__setitem__, fnn, 1)
        assert_raises(ValueError, b.__getitem__, fnn)
        b[0][fn1] = 2
        assert_equal(b[fn1], 2)
        # Subfield
        assert_raises(ValueError, b[0].__setitem__, fnn, 1)
        assert_raises(ValueError, b[0].__getitem__, fnn)
        # Subfield
        fn3 = 'f3'
        sfn1 = 'sf1'
        b[fn3][sfn1] = 1
        assert_equal(b[fn3][sfn1], 1)
        assert_raises(ValueError, b[fn3].__setitem__, fnn, 1)
        assert_raises(ValueError, b[fn3].__getitem__, fnn)
        # multiple subfields
        fn2 = 'f2'
        b[fn2] = 3

        assert_equal(b[['f1', 'f2']][0].tolist(), (2, 3))
        assert_equal(b[['f2', 'f1']][0].tolist(), (3, 2))
        assert_equal(b[['f1', 'f3']][0].tolist(), (2, (1,)))

        # non-ascii unicode field indexing is well behaved
        assert_raises(ValueError, a.__setitem__, '\u03e0', 1)
        assert_raises(ValueError, a.__getitem__, '\u03e0')

    def test_record_hash(self):
        a = np.array([(1, 2), (1, 2)], dtype='i1,i2')
        a.flags.writeable = False
        b = np.array([(1, 2), (3, 4)], dtype=[('num1', 'i1'), ('num2', 'i2')])
        b.flags.writeable = False
        c = np.array([(1, 2), (3, 4)], dtype='i1,i2')
        c.flags.writeable = False
        assert_(hash(a[0]) == hash(a[1]))
        assert_(hash(a[0]) == hash(b[0]))
        assert_(hash(a[0]) != hash(b[1]))
        assert_(hash(c[0]) == hash(a[0]) and c[0] == a[0])

    def test_record_no_hash(self):
        a = np.array([(1, 2), (1, 2)], dtype='i1,i2')
        assert_raises(TypeError, hash, a[0])

    def test_empty_structure_creation(self):
        # make sure these do not raise errors (gh-5631)
        np.array([()], dtype={'names': [], 'formats': [],
                           'offsets': [], 'itemsize': 12})
        np.array([(), (), (), (), ()], dtype={'names': [], 'formats': [],
                                           'offsets': [], 'itemsize': 12})

    def test_multifield_indexing_view(self):
        a = np.ones(3, dtype=[('a', 'i4'), ('b', 'f4'), ('c', 'u4')])
        v = a[['a', 'c']]
        assert_(v.base is a)
        assert_(v.dtype == np.dtype({'names': ['a', 'c'],
                                     'formats': ['i4', 'u4'],
                                     'offsets': [0, 8]}))
        v[:] = (4, 5)
        assert_equal(a[0].item(), (4, 1, 5))

class TestView:
    def test_basic(self):
        x = np.array([(1, 2, 3, 4), (5, 6, 7, 8)],
                     dtype=[('r', np.int8), ('g', np.int8),
                            ('b', np.int8), ('a', np.int8)])
        # We must be specific about the endianness here:
        y = x.view(dtype='<i4')
        # ... and again without the keyword.
        z = x.view('<i4')
        assert_array_equal(y, z)
        assert_array_equal(y, [67305985, 134678021])


def _mean(a, **args):
    return a.mean(**args)


def _var(a, **args):
    return a.var(**args)


def _std(a, **args):
    return a.std(**args)


class TestStats:

    funcs = [_mean, _var, _std]

    def setup_method(self):
        np.random.seed(range(3))
        self.rmat = np.random.random((4, 5))
        self.cmat = self.rmat + 1j * self.rmat
        self.omat = np.array([Decimal(str(r)) for r in self.rmat.flat])
        self.omat = self.omat.reshape(4, 5)

    def test_python_type(self):
        for x in (np.float16(1.), 1, 1., 1 + 0j):
            assert_equal(np.mean([x]), 1.)
            assert_equal(np.std([x]), 0.)
            assert_equal(np.var([x]), 0.)

    def test_keepdims(self):
        mat = np.eye(3)
        for f in self.funcs:
            for axis in [0, 1]:
                res = f(mat, axis=axis, keepdims=True)
                assert_(res.ndim == mat.ndim)
                assert_(res.shape[axis] == 1)
            for axis in [None]:
                res = f(mat, axis=axis, keepdims=True)
                assert_(res.shape == (1, 1))

    def test_out(self):
        mat = np.eye(3)
        for f in self.funcs:
            out = np.zeros(3)
            tgt = f(mat, axis=1)
            res = f(mat, axis=1, out=out)
            assert_almost_equal(res, out)
            assert_almost_equal(res, tgt)
        out = np.empty(2)
        assert_raises(ValueError, f, mat, axis=1, out=out)
        out = np.empty((2, 2))
        assert_raises(ValueError, f, mat, axis=1, out=out)

    def test_dtype_from_input(self):

        icodes = np.typecodes['AllInteger']
        fcodes = np.typecodes['AllFloat']

        # object type
        for f in self.funcs:
            mat = np.array([[Decimal(1)] * 3] * 3)
            tgt = mat.dtype.type
            res = f(mat, axis=1).dtype.type
            assert_(res is tgt)
            # scalar case
            res = type(f(mat, axis=None))
            assert_(res is Decimal)

        # integer types
        for f in self.funcs:
            for c in icodes:
                mat = np.eye(3, dtype=c)
                tgt = np.float64
                res = f(mat, axis=1).dtype.type
                assert_(res is tgt)
                # scalar case
                res = f(mat, axis=None).dtype.type
                assert_(res is tgt)

        # mean for float types
        for f in [_mean]:
            for c in fcodes:
                mat = np.eye(3, dtype=c)
                tgt = mat.dtype.type
                res = f(mat, axis=1).dtype.type
                assert_(res is tgt)
                # scalar case
                res = f(mat, axis=None).dtype.type
                assert_(res is tgt)

        # var, std for float types
        for f in [_var, _std]:
            for c in fcodes:
                mat = np.eye(3, dtype=c)
                # deal with complex types
                tgt = mat.real.dtype.type
                res = f(mat, axis=1).dtype.type
                assert_(res is tgt)
                # scalar case
                res = f(mat, axis=None).dtype.type
                assert_(res is tgt)

    def test_dtype_from_dtype(self):
        mat = np.eye(3)

        # stats for integer types
        # FIXME:
        # this needs definition as there are lots places along the line
        # where type casting may take place.

        # for f in self.funcs:
        #    for c in np.typecodes['AllInteger']:
        #        tgt = np.dtype(c).type
        #        res = f(mat, axis=1, dtype=c).dtype.type
        #        assert_(res is tgt)
        #        # scalar case
        #        res = f(mat, axis=None, dtype=c).dtype.type
        #        assert_(res is tgt)

        # stats for float types
        for f in self.funcs:
            for c in np.typecodes['AllFloat']:
                tgt = np.dtype(c).type
                res = f(mat, axis=1, dtype=c).dtype.type
                assert_(res is tgt)
                # scalar case
                res = f(mat, axis=None, dtype=c).dtype.type
                assert_(res is tgt)

    def test_ddof(self):
        for f in [_var]:
            for ddof in range(3):
                dim = self.rmat.shape[1]
                tgt = f(self.rmat, axis=1) * dim
                res = f(self.rmat, axis=1, ddof=ddof) * (dim - ddof)
        for f in [_std]:
            for ddof in range(3):
                dim = self.rmat.shape[1]
                tgt = f(self.rmat, axis=1) * np.sqrt(dim)
                res = f(self.rmat, axis=1, ddof=ddof) * np.sqrt(dim - ddof)
                assert_almost_equal(res, tgt)
                assert_almost_equal(res, tgt)

    def test_ddof_too_big(self):
        dim = self.rmat.shape[1]
        for f in [_var, _std]:
            for ddof in range(dim, dim + 2):
                with warnings.catch_warnings(record=True) as w:
                    warnings.simplefilter('always')
                    res = f(self.rmat, axis=1, ddof=ddof)
                    assert_(not (res < 0).any())
                    assert_(len(w) > 0)
                    assert_(issubclass(w[0].category, RuntimeWarning))

    def test_empty(self):
        A = np.zeros((0, 3))
        for f in self.funcs:
            for axis in [0, None]:
                with warnings.catch_warnings(record=True) as w:
                    warnings.simplefilter('always')
                    assert_(np.isnan(f(A, axis=axis)).all())
                    assert_(len(w) > 0)
                    assert_(issubclass(w[0].category, RuntimeWarning))
            for axis in [1]:
                with warnings.catch_warnings(record=True) as w:
                    warnings.simplefilter('always')
                    assert_equal(f(A, axis=axis), np.zeros([]))

    def test_mean_values(self):
        for mat in [self.rmat, self.cmat, self.omat]:
            for axis in [0, 1]:
                tgt = mat.sum(axis=axis)
                res = _mean(mat, axis=axis) * mat.shape[axis]
                assert_almost_equal(res, tgt)
            for axis in [None]:
                tgt = mat.sum(axis=axis)
                res = _mean(mat, axis=axis) * np.prod(mat.shape)
                assert_almost_equal(res, tgt)

    def test_mean_float16(self):
        # This fail if the sum inside mean is done in float16 instead
        # of float32.
        assert_(_mean(np.ones(100000, dtype='float16')) == 1)

    def test_mean_axis_error(self):
        # Ensure that AxisError is raised instead of IndexError when axis is
        # out of bounds, see gh-15817.
        with assert_raises(np.exceptions.AxisError):
            np.arange(10).mean(axis=2)

    def test_mean_where(self):
        a = np.arange(16).reshape((4, 4))
        wh_full = np.array([[False, True, False, True],
                            [True, False, True, False],
                            [True, True, False, False],
                            [False, False, True, True]])
        wh_partial = np.array([[False],
                               [True],
                               [True],
                               [False]])
        _cases = [(1, True, [1.5, 5.5, 9.5, 13.5]),
                  (0, wh_full, [6., 5., 10., 9.]),
                  (1, wh_full, [2., 5., 8.5, 14.5]),
                  (0, wh_partial, [6., 7., 8., 9.])]
        for _ax, _wh, _res in _cases:
            assert_allclose(a.mean(axis=_ax, where=_wh),
                            np.array(_res))
            assert_allclose(np.mean(a, axis=_ax, where=_wh),
                            np.array(_res))

        a3d = np.arange(16).reshape((2, 2, 4))
        _wh_partial = np.array([False, True, True, False])
        _res = [[1.5, 5.5], [9.5, 13.5]]
        assert_allclose(a3d.mean(axis=2, where=_wh_partial),
                        np.array(_res))
        assert_allclose(np.mean(a3d, axis=2, where=_wh_partial),
                        np.array(_res))

        with pytest.warns(RuntimeWarning) as w:
            assert_allclose(a.mean(axis=1, where=wh_partial),
                            np.array([np.nan, 5.5, 9.5, np.nan]))
        with pytest.warns(RuntimeWarning) as w:
            assert_equal(a.mean(where=False), np.nan)
        with pytest.warns(RuntimeWarning) as w:
            assert_equal(np.mean(a, where=False), np.nan)

    def test_var_values(self):
        for mat in [self.rmat, self.cmat, self.omat]:
            for axis in [0, 1, None]:
                msqr = _mean(mat * mat.conj(), axis=axis)
                mean = _mean(mat, axis=axis)
                tgt = msqr - mean * mean.conjugate()
                res = _var(mat, axis=axis)
                assert_almost_equal(res, tgt)

    @pytest.mark.parametrize(('complex_dtype', 'ndec'), (
        ('complex64', 6),
        ('complex128', 7),
        ('clongdouble', 7),
    ))
    def test_var_complex_values(self, complex_dtype, ndec):
        # Test fast-paths for every builtin complex type
        for axis in [0, 1, None]:
            mat = self.cmat.copy().astype(complex_dtype)
            msqr = _mean(mat * mat.conj(), axis=axis)
            mean = _mean(mat, axis=axis)
            tgt = msqr - mean * mean.conjugate()
            res = _var(mat, axis=axis)
            assert_almost_equal(res, tgt, decimal=ndec)

    def test_var_dimensions(self):
        # _var paths for complex number introduce additions on views that
        # increase dimensions. Ensure this generalizes to higher dims
        mat = np.stack([self.cmat] * 3)
        for axis in [0, 1, 2, -1, None]:
            msqr = _mean(mat * mat.conj(), axis=axis)
            mean = _mean(mat, axis=axis)
            tgt = msqr - mean * mean.conjugate()
            res = _var(mat, axis=axis)
            assert_almost_equal(res, tgt)

    def test_var_complex_byteorder(self):
        # Test that var fast-path does not cause failures for complex arrays
        # with non-native byteorder
        cmat = self.cmat.copy().astype('complex128')
        cmat_swapped = cmat.astype(cmat.dtype.newbyteorder())
        assert_almost_equal(cmat.var(), cmat_swapped.var())

    def test_var_axis_error(self):
        # Ensure that AxisError is raised instead of IndexError when axis is
        # out of bounds, see gh-15817.
        with assert_raises(np.exceptions.AxisError):
            np.arange(10).var(axis=2)

    def test_var_where(self):
        a = np.arange(25).reshape((5, 5))
        wh_full = np.array([[False, True, False, True, True],
                            [True, False, True, True, False],
                            [True, True, False, False, True],
                            [False, True, True, False, True],
                            [True, False, True, True, False]])
        wh_partial = np.array([[False],
                               [True],
                               [True],
                               [False],
                               [True]])
        _cases = [(0, True, [50., 50., 50., 50., 50.]),
                  (1, True, [2., 2., 2., 2., 2.])]
        for _ax, _wh, _res in _cases:
            assert_allclose(a.var(axis=_ax, where=_wh),
                            np.array(_res))
            assert_allclose(np.var(a, axis=_ax, where=_wh),
                            np.array(_res))

        a3d = np.arange(16).reshape((2, 2, 4))
        _wh_partial = np.array([False, True, True, False])
        _res = [[0.25, 0.25], [0.25, 0.25]]
        assert_allclose(a3d.var(axis=2, where=_wh_partial),
                        np.array(_res))
        assert_allclose(np.var(a3d, axis=2, where=_wh_partial),
                        np.array(_res))

        assert_allclose(np.var(a, axis=1, where=wh_full),
                        np.var(a[wh_full].reshape((5, 3)), axis=1))
        assert_allclose(np.var(a, axis=0, where=wh_partial),
                        np.var(a[wh_partial[:, 0]], axis=0))
        with pytest.warns(RuntimeWarning) as w:
            assert_equal(a.var(where=False), np.nan)
        with pytest.warns(RuntimeWarning) as w:
            assert_equal(np.var(a, where=False), np.nan)

    def test_std_values(self):
        for mat in [self.rmat, self.cmat, self.omat]:
            for axis in [0, 1, None]:
                tgt = np.sqrt(_var(mat, axis=axis))
                res = _std(mat, axis=axis)
                assert_almost_equal(res, tgt)

    def test_std_where(self):
        a = np.arange(25).reshape((5, 5))[::-1]
        whf = np.array([[False, True, False, True, True],
                        [True, False, True, False, True],
                        [True, True, False, True, False],
                        [True, False, True, True, False],
                        [False, True, False, True, True]])
        whp = np.array([[False],
                        [False],
                        [True],
                        [True],
                        [False]])
        _cases = [
            (0, True, 7.07106781 * np.ones(5)),
            (1, True, 1.41421356 * np.ones(5)),
            (0, whf,
             np.array([4.0824829, 8.16496581, 5., 7.39509973, 8.49836586])),
            (0, whp, 2.5 * np.ones(5))
        ]
        for _ax, _wh, _res in _cases:
            assert_allclose(a.std(axis=_ax, where=_wh), _res)
            assert_allclose(np.std(a, axis=_ax, where=_wh), _res)

        a3d = np.arange(16).reshape((2, 2, 4))
        _wh_partial = np.array([False, True, True, False])
        _res = [[0.5, 0.5], [0.5, 0.5]]
        assert_allclose(a3d.std(axis=2, where=_wh_partial),
                        np.array(_res))
        assert_allclose(np.std(a3d, axis=2, where=_wh_partial),
                        np.array(_res))

        assert_allclose(a.std(axis=1, where=whf),
                        np.std(a[whf].reshape((5, 3)), axis=1))
        assert_allclose(np.std(a, axis=1, where=whf),
                        (a[whf].reshape((5, 3))).std(axis=1))
        assert_allclose(a.std(axis=0, where=whp),
                        np.std(a[whp[:, 0]], axis=0))
        assert_allclose(np.std(a, axis=0, where=whp),
                        (a[whp[:, 0]]).std(axis=0))
        with pytest.warns(RuntimeWarning) as w:
            assert_equal(a.std(where=False), np.nan)
        with pytest.warns(RuntimeWarning) as w:
            assert_equal(np.std(a, where=False), np.nan)

    def test_subclass(self):
        class TestArray(np.ndarray):
            def __new__(cls, data, info):
                result = np.array(data)
                result = result.view(cls)
                result.info = info
                return result

            def __array_finalize__(self, obj):
                self.info = getattr(obj, "info", '')

        dat = TestArray([[1, 2, 3, 4], [5, 6, 7, 8]], 'jubba')
        res = dat.mean(1)
        assert_(res.info == dat.info)
        res = dat.std(1)
        assert_(res.info == dat.info)
        res = dat.var(1)
        assert_(res.info == dat.info)


class TestVdot:
    def test_basic(self):
        dt_numeric = np.typecodes['AllFloat'] + np.typecodes['AllInteger']
        dt_complex = np.typecodes['Complex']

        # test real
        a = np.eye(3)
        for dt in dt_numeric + 'O':
            b = a.astype(dt)
            res = np.vdot(b, b)
            assert_(np.isscalar(res))
            assert_equal(np.vdot(b, b), 3)

        # test complex
        a = np.eye(3) * 1j
        for dt in dt_complex + 'O':
            b = a.astype(dt)
            res = np.vdot(b, b)
            assert_(np.isscalar(res))
            assert_equal(np.vdot(b, b), 3)

        # test boolean
        b = np.eye(3, dtype=bool)
        res = np.vdot(b, b)
        assert_(np.isscalar(res))
        assert_equal(np.vdot(b, b), True)

    def test_vdot_array_order(self):
        a = np.array([[1, 2], [3, 4]], order='C')
        b = np.array([[1, 2], [3, 4]], order='F')
        res = np.vdot(a, a)

        # integer arrays are exact
        assert_equal(np.vdot(a, b), res)
        assert_equal(np.vdot(b, a), res)
        assert_equal(np.vdot(b, b), res)

    def test_vdot_uncontiguous(self):
        for size in [2, 1000]:
            # Different sizes match different branches in vdot.
            a = np.zeros((size, 2, 2))
            b = np.zeros((size, 2, 2))
            a[:, 0, 0] = np.arange(size)
            b[:, 0, 0] = np.arange(size) + 1
            # Make a and b uncontiguous:
            a = a[..., 0]
            b = b[..., 0]

            assert_equal(np.vdot(a, b),
                         np.vdot(a.flatten(), b.flatten()))
            assert_equal(np.vdot(a, b.copy()),
                         np.vdot(a.flatten(), b.flatten()))
            assert_equal(np.vdot(a.copy(), b),
                         np.vdot(a.flatten(), b.flatten()))
            assert_equal(np.vdot(a.copy('F'), b),
                         np.vdot(a.flatten(), b.flatten()))
            assert_equal(np.vdot(a, b.copy('F')),
                         np.vdot(a.flatten(), b.flatten()))


class TestDot:
    def setup_method(self):
        np.random.seed(128)
        self.A = np.random.rand(4, 2)
        self.b1 = np.random.rand(2, 1)
        self.b2 = np.random.rand(2)
        self.b3 = np.random.rand(1, 2)
        self.b4 = np.random.rand(4)
        self.N = 7

    def test_dotmatmat(self):
        A = self.A
        res = np.dot(A.transpose(), A)
        tgt = np.array([[1.45046013, 0.86323640],
                        [0.86323640, 0.84934569]])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotmatvec(self):
        A, b1 = self.A, self.b1
        res = np.dot(A, b1)
        tgt = np.array([[0.32114320], [0.04889721],
                        [0.15696029], [0.33612621]])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotmatvec2(self):
        A, b2 = self.A, self.b2
        res = np.dot(A, b2)
        tgt = np.array([0.29677940, 0.04518649, 0.14468333, 0.31039293])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotvecmat(self):
        A, b4 = self.A, self.b4
        res = np.dot(b4, A)
        tgt = np.array([1.23495091, 1.12222648])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotvecmat2(self):
        b3, A = self.b3, self.A
        res = np.dot(b3, A.transpose())
        tgt = np.array([[0.58793804, 0.08957460, 0.30605758, 0.62716383]])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotvecmat3(self):
        A, b4 = self.A, self.b4
        res = np.dot(A.transpose(), b4)
        tgt = np.array([1.23495091, 1.12222648])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotvecvecouter(self):
        b1, b3 = self.b1, self.b3
        res = np.dot(b1, b3)
        tgt = np.array([[0.20128610, 0.08400440], [0.07190947, 0.03001058]])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotvecvecinner(self):
        b1, b3 = self.b1, self.b3
        res = np.dot(b3, b1)
        tgt = np.array([[0.23129668]])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotcolumnvect1(self):
        b1 = np.ones((3, 1))
        b2 = [5.3]
        res = np.dot(b1, b2)
        tgt = np.array([5.3, 5.3, 5.3])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotcolumnvect2(self):
        b1 = np.ones((3, 1)).transpose()
        b2 = [6.2]
        res = np.dot(b2, b1)
        tgt = np.array([6.2, 6.2, 6.2])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotvecscalar(self):
        np.random.seed(100)
        b1 = np.random.rand(1, 1)
        b2 = np.random.rand(1, 4)
        res = np.dot(b1, b2)
        tgt = np.array([[0.15126730, 0.23068496, 0.45905553, 0.00256425]])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_dotvecscalar2(self):
        np.random.seed(100)
        b1 = np.random.rand(4, 1)
        b2 = np.random.rand(1, 1)
        res = np.dot(b1, b2)
        tgt = np.array([[0.00256425], [0.00131359], [0.00200324], [0.00398638]])
        assert_almost_equal(res, tgt, decimal=self.N)

    def test_all(self):
        dims = [(), (1,), (1, 1)]
        dout = [(), (1,), (1, 1), (1,), (), (1,), (1, 1), (1,), (1, 1)]
        for dim, (dim1, dim2) in zip(dout, itertools.product(dims, dims)):
            b1 = np.zeros(dim1)
            b2 = np.zeros(dim2)
            res = np.dot(b1, b2)
            tgt = np.zeros(dim)
            assert_(res.shape == tgt.shape)
            assert_almost_equal(res, tgt, decimal=self.N)

    def test_vecobject(self):
        class Vec:
            def __init__(self, sequence=None):
                if sequence is None:
                    sequence = []
                self.array = np.array(sequence)

            def __add__(self, other):
                out = Vec()
                out.array = self.array + other.array
                return out

            def __sub__(self, other):
                out = Vec()
                out.array = self.array - other.array
                return out

            def __mul__(self, other):  # with scalar
                out = Vec(self.array.copy())
                out.array *= other
                return out

            def __rmul__(self, other):
                return self * other

        U_non_cont = np.transpose([[1., 1.], [1., 2.]])
        U_cont = np.ascontiguousarray(U_non_cont)
        x = np.array([Vec([1., 0.]), Vec([0., 1.])])
        zeros = np.array([Vec([0., 0.]), Vec([0., 0.])])
        zeros_test = np.dot(U_cont, x) - np.dot(U_non_cont, x)
        assert_equal(zeros[0].array, zeros_test[0].array)
        assert_equal(zeros[1].array, zeros_test[1].array)

    def test_dot_2args(self):

        a = np.array([[1, 2], [3, 4]], dtype=float)
        b = np.array([[1, 0], [1, 1]], dtype=float)
        c = np.array([[3, 2], [7, 4]], dtype=float)

        d = dot(a, b)
        assert_allclose(c, d)

    def test_dot_3args(self):

        np.random.seed(22)
        f = np.random.random_sample((1024, 16))
        v = np.random.random_sample((16, 32))

        r = np.empty((1024, 32))
        if HAS_REFCOUNT:
            orig_refcount = sys.getrefcount(r)
        for i in range(12):
            dot(f, v, r)
        if HAS_REFCOUNT:
            assert_equal(sys.getrefcount(r), orig_refcount)
        r2 = dot(f, v, out=None)
        assert_array_equal(r2, r)
        assert_(r is dot(f, v, out=r))

        v = v[:, 0].copy()  # v.shape == (16,)
        r = r[:, 0].copy()  # r.shape == (1024,)
        r2 = dot(f, v)
        assert_(r is dot(f, v, r))
        assert_array_equal(r2, r)

    def test_dot_3args_errors(self):

        np.random.seed(22)
        f = np.random.random_sample((1024, 16))
        v = np.random.random_sample((16, 32))

        r = np.empty((1024, 31))
        assert_raises(ValueError, dot, f, v, r)

        r = np.empty((1024,))
        assert_raises(ValueError, dot, f, v, r)

        r = np.empty((32,))
        assert_raises(ValueError, dot, f, v, r)

        r = np.empty((32, 1024))
        assert_raises(ValueError, dot, f, v, r)
        assert_raises(ValueError, dot, f, v, r.T)

        r = np.empty((1024, 64))
        assert_raises(ValueError, dot, f, v, r[:, ::2])
        assert_raises(ValueError, dot, f, v, r[:, :32])

        r = np.empty((1024, 32), dtype=np.float32)
        assert_raises(ValueError, dot, f, v, r)

        r = np.empty((1024, 32), dtype=int)
        assert_raises(ValueError, dot, f, v, r)

    def test_dot_out_result(self):
        x = np.ones((), dtype=np.float16)
        y = np.ones((5,), dtype=np.float16)
        z = np.zeros((5,), dtype=np.float16)
        res = x.dot(y, out=z)
        assert np.array_equal(res, y)
        assert np.array_equal(z, y)

    def test_dot_out_aliasing(self):
        x = np.ones((), dtype=np.float16)
        y = np.ones((5,), dtype=np.float16)
        z = np.zeros((5,), dtype=np.float16)
        res = x.dot(y, out=z)
        z[0] = 2
        assert np.array_equal(res, z)

    def test_dot_array_order(self):
        a = np.array([[1, 2], [3, 4]], order='C')
        b = np.array([[1, 2], [3, 4]], order='F')
        res = np.dot(a, a)

        # integer arrays are exact
        assert_equal(np.dot(a, b), res)
        assert_equal(np.dot(b, a), res)
        assert_equal(np.dot(b, b), res)

    def test_accelerate_framework_sgemv_fix(self):

        def aligned_array(shape, align, dtype, order='C'):
            d = dtype(0)
            N = np.prod(shape)
            tmp = np.zeros(N * d.nbytes + align, dtype=np.uint8)
            address = tmp.__array_interface__["data"][0]
            for offset in range(align):
                if (address + offset) % align == 0:
                    break
            tmp = tmp[offset:offset + N * d.nbytes].view(dtype=dtype)
            return tmp.reshape(shape, order=order)

        def as_aligned(arr, align, dtype, order='C'):
            aligned = aligned_array(arr.shape, align, dtype, order)
            aligned[:] = arr[:]
            return aligned

        def assert_dot_close(A, X, desired):
            assert_allclose(np.dot(A, X), desired, rtol=1e-5, atol=1e-7)

        m = aligned_array(100, 15, np.float32)
        s = aligned_array((100, 100), 15, np.float32)
        np.dot(s, m)  # this will always segfault if the bug is present

        testdata = itertools.product((15, 32), (10000,), (200, 89), ('C', 'F'))
        for align, m, n, a_order in testdata:
            # Calculation in double precision
            A_d = np.random.rand(m, n)
            X_d = np.random.rand(n)
            desired = np.dot(A_d, X_d)
            # Calculation with aligned single precision
            A_f = as_aligned(A_d, align, np.float32, order=a_order)
            X_f = as_aligned(X_d, align, np.float32)
            assert_dot_close(A_f, X_f, desired)
            # Strided A rows
            A_d_2 = A_d[::2]
            desired = np.dot(A_d_2, X_d)
            A_f_2 = A_f[::2]
            assert_dot_close(A_f_2, X_f, desired)
            # Strided A columns, strided X vector
            A_d_22 = A_d_2[:, ::2]
            X_d_2 = X_d[::2]
            desired = np.dot(A_d_22, X_d_2)
            A_f_22 = A_f_2[:, ::2]
            X_f_2 = X_f[::2]
            assert_dot_close(A_f_22, X_f_2, desired)
            # Check the strides are as expected
            if a_order == 'F':
                assert_equal(A_f_22.strides, (8, 8 * m))
            else:
                assert_equal(A_f_22.strides, (8 * n, 8))
            assert_equal(X_f_2.strides, (8,))
            # Strides in A rows + cols only
            X_f_2c = as_aligned(X_f_2, align, np.float32)
            assert_dot_close(A_f_22, X_f_2c, desired)
            # Strides just in A cols
            A_d_12 = A_d[:, ::2]
            desired = np.dot(A_d_12, X_d_2)
            A_f_12 = A_f[:, ::2]
            assert_dot_close(A_f_12, X_f_2c, desired)
            # Strides in A cols and X
            assert_dot_close(A_f_12, X_f_2, desired)

    @pytest.mark.slow
    @pytest.mark.parametrize("dtype", [np.float64, np.complex128])
    @requires_memory(free_bytes=18e9)  # complex case needs 18GiB+
    def test_huge_vectordot(self, dtype):
        # Large vector multiplications are chunked with 32bit BLAS
        # Test that the chunking does the right thing, see also gh-22262
        data = np.ones(2**30 + 100, dtype=dtype)
        res = np.dot(data, data)
        assert res == 2**30 + 100

    def test_dtype_discovery_fails(self):
        # See gh-14247, error checking was missing for failed dtype discovery
        class BadObject:
            def __array__(self, dtype=None, copy=None):
                raise TypeError("just this tiny mint leaf")

        with pytest.raises(TypeError):
            np.dot(BadObject(), BadObject())

        with pytest.raises(TypeError):
            np.dot(3.0, BadObject())


class MatmulCommon:
    """Common tests for '@' operator and numpy.matmul.

    """
    # Should work with these types. Will want to add
    # "O" at some point
    types = "?bhilqBHILQefdgFDGO"

    def test_exceptions(self):
        dims = [
            ((1,), (2,)),            # mismatched vector vector
            ((2, 1,), (2,)),         # mismatched matrix vector
            ((2,), (1, 2)),          # mismatched vector matrix
            ((1, 2), (3, 1)),        # mismatched matrix matrix
            ((1,), ()),              # vector scalar
            ((), (1)),               # scalar vector
            ((1, 1), ()),            # matrix scalar
            ((), (1, 1)),            # scalar matrix
            ((2, 2, 1), (3, 1, 2)),  # cannot broadcast
            ]

        for dt, (dm1, dm2) in itertools.product(self.types, dims):
            a = np.ones(dm1, dtype=dt)
            b = np.ones(dm2, dtype=dt)
            assert_raises(ValueError, self.matmul, a, b)

    def test_shapes(self):
        dims = [
            ((1, 1), (2, 1, 1)),     # broadcast first argument
            ((2, 1, 1), (1, 1)),     # broadcast second argument
            ((2, 1, 1), (2, 1, 1)),  # matrix stack sizes match
            ]

        for dt, (dm1, dm2) in itertools.product(self.types, dims):
            a = np.ones(dm1, dtype=dt)
            b = np.ones(dm2, dtype=dt)
            res = self.matmul(a, b)
            assert_(res.shape == (2, 1, 1))

        # vector vector returns scalars.
        for dt in self.types:
            a = np.ones((2,), dtype=dt)
            b = np.ones((2,), dtype=dt)
            c = self.matmul(a, b)
            assert_(np.array(c).shape == ())

    def test_result_types(self):
        mat = np.ones((1, 1))
        vec = np.ones((1,))
        for dt in self.types:
            m = mat.astype(dt)
            v = vec.astype(dt)
            for arg in [(m, v), (v, m), (m, m)]:
                res = self.matmul(*arg)
                assert_(res.dtype == dt)

            # vector vector returns scalars
            if dt != "O":
                res = self.matmul(v, v)
                assert_(type(res) is np.dtype(dt).type)

    def test_scalar_output(self):
        vec1 = np.array([2])
        vec2 = np.array([3, 4]).reshape(1, -1)
        tgt = np.array([6, 8])
        for dt in self.types[1:]:
            v1 = vec1.astype(dt)
            v2 = vec2.astype(dt)
            res = self.matmul(v1, v2)
            assert_equal(res, tgt)
            res = self.matmul(v2.T, v1)
            assert_equal(res, tgt)

        # boolean type
        vec = np.array([True, True], dtype='?').reshape(1, -1)
        res = self.matmul(vec[:, 0], vec)
        assert_equal(res, True)

    def test_vector_vector_values(self):
        vec1 = np.array([1, 2])
        vec2 = np.array([3, 4]).reshape(-1, 1)
        tgt1 = np.array([11])
        tgt2 = np.array([[3, 6], [4, 8]])
        for dt in self.types[1:]:
            v1 = vec1.astype(dt)
            v2 = vec2.astype(dt)
            res = self.matmul(v1, v2)
            assert_equal(res, tgt1)
            # no broadcast, we must make v1 into a 2d ndarray
            res = self.matmul(v2, v1.reshape(1, -1))
            assert_equal(res, tgt2)

        # boolean type
        vec = np.array([True, True], dtype='?')
        res = self.matmul(vec, vec)
        assert_equal(res, True)

    def test_vector_matrix_values(self):
        vec = np.array([1, 2])
        mat1 = np.array([[1, 2], [3, 4]])
        mat2 = np.stack([mat1] * 2, axis=0)
        tgt1 = np.array([7, 10])
        tgt2 = np.stack([tgt1] * 2, axis=0)
        for dt in self.types[1:]:
            v = vec.astype(dt)
            m1 = mat1.astype(dt)
            m2 = mat2.astype(dt)
            res = self.matmul(v, m1)
            assert_equal(res, tgt1)
            res = self.matmul(v, m2)
            assert_equal(res, tgt2)

        # boolean type
        vec = np.array([True, False])
        mat1 = np.array([[True, False], [False, True]])
        mat2 = np.stack([mat1] * 2, axis=0)
        tgt1 = np.array([True, False])
        tgt2 = np.stack([tgt1] * 2, axis=0)

        res = self.matmul(vec, mat1)
        assert_equal(res, tgt1)
        res = self.matmul(vec, mat2)
        assert_equal(res, tgt2)

    def test_matrix_vector_values(self):
        vec = np.array([1, 2])
        mat1 = np.array([[1, 2], [3, 4]])
        mat2 = np.stack([mat1] * 2, axis=0)
        tgt1 = np.array([5, 11])
        tgt2 = np.stack([tgt1] * 2, axis=0)
        for dt in self.types[1:]:
            v = vec.astype(dt)
            m1 = mat1.astype(dt)
            m2 = mat2.astype(dt)
            res = self.matmul(m1, v)
            assert_equal(res, tgt1)
            res = self.matmul(m2, v)
            assert_equal(res, tgt2)

        # boolean type
        vec = np.array([True, False])
        mat1 = np.array([[True, False], [False, True]])
        mat2 = np.stack([mat1] * 2, axis=0)
        tgt1 = np.array([True, False])
        tgt2 = np.stack([tgt1] * 2, axis=0)

        res = self.matmul(vec, mat1)
        assert_equal(res, tgt1)
        res = self.matmul(vec, mat2)
        assert_equal(res, tgt2)

    def test_matrix_matrix_values(self):
        mat1 = np.array([[1, 2], [3, 4]])
        mat2 = np.array([[1, 0], [1, 1]])
        mat12 = np.stack([mat1, mat2], axis=0)
        mat21 = np.stack([mat2, mat1], axis=0)
        tgt11 = np.array([[7, 10], [15, 22]])
        tgt12 = np.array([[3, 2], [7, 4]])
        tgt21 = np.array([[1, 2], [4, 6]])
        tgt12_21 = np.stack([tgt12, tgt21], axis=0)
        tgt11_12 = np.stack((tgt11, tgt12), axis=0)
        tgt11_21 = np.stack((tgt11, tgt21), axis=0)
        for dt in self.types[1:]:
            m1 = mat1.astype(dt)
            m2 = mat2.astype(dt)
            m12 = mat12.astype(dt)
            m21 = mat21.astype(dt)

            # matrix @ matrix
            res = self.matmul(m1, m2)
            assert_equal(res, tgt12)
            res = self.matmul(m2, m1)
            assert_equal(res, tgt21)

            # stacked @ matrix
            res = self.matmul(m12, m1)
            assert_equal(res, tgt11_21)

            # matrix @ stacked
            res = self.matmul(m1, m12)
            assert_equal(res, tgt11_12)

            # stacked @ stacked
            res = self.matmul(m12, m21)
            assert_equal(res, tgt12_21)

        # boolean type
        m1 = np.array([[1, 1], [0, 0]], dtype=np.bool)
        m2 = np.array([[1, 0], [1, 1]], dtype=np.bool)
        m12 = np.stack([m1, m2], axis=0)
        m21 = np.stack([m2, m1], axis=0)
        tgt11 = m1
        tgt12 = m1
        tgt21 = np.array([[1, 1], [1, 1]], dtype=np.bool)
        tgt12_21 = np.stack([tgt12, tgt21], axis=0)
        tgt11_12 = np.stack((tgt11, tgt12), axis=0)
        tgt11_21 = np.stack((tgt11, tgt21), axis=0)

        # matrix @ matrix
        res = self.matmul(m1, m2)
        assert_equal(res, tgt12)
        res = self.matmul(m2, m1)
        assert_equal(res, tgt21)

        # stacked @ matrix
        res = self.matmul(m12, m1)
        assert_equal(res, tgt11_21)

        # matrix @ stacked
        res = self.matmul(m1, m12)
        assert_equal(res, tgt11_12)

        # stacked @ stacked
        res = self.matmul(m12, m21)
        assert_equal(res, tgt12_21)


class TestMatmul(MatmulCommon):
    matmul = np.matmul

    def test_out_arg(self):
        a = np.ones((5, 2), dtype=float)
        b = np.array([[1, 3], [5, 7]], dtype=float)
        tgt = np.dot(a, b)

        # test as positional argument
        msg = "out positional argument"
        out = np.zeros((5, 2), dtype=float)
        self.matmul(a, b, out)
        assert_array_equal(out, tgt, err_msg=msg)

        # test as keyword argument
        msg = "out keyword argument"
        out = np.zeros((5, 2), dtype=float)
        self.matmul(a, b, out=out)
        assert_array_equal(out, tgt, err_msg=msg)

        # test out with not allowed type cast (safe casting)
        msg = "Cannot cast ufunc .* output"
        out = np.zeros((5, 2), dtype=np.int32)
        assert_raises_regex(TypeError, msg, self.matmul, a, b, out=out)

        # test out with type upcast to complex
        out = np.zeros((5, 2), dtype=np.complex128)
        c = self.matmul(a, b, out=out)
        assert_(c is out)
        with suppress_warnings() as sup:
            sup.filter(ComplexWarning, '')
            c = c.astype(tgt.dtype)
        assert_array_equal(c, tgt)

    def test_empty_out(self):
        # Check that the output cannot be broadcast, so that it cannot be
        # size zero when the outer dimensions (iterator size) has size zero.
        arr = np.ones((0, 1, 1))
        out = np.ones((1, 1, 1))
        assert self.matmul(arr, arr).shape == (0, 1, 1)

        with pytest.raises(ValueError, match=r"non-broadcastable"):
            self.matmul(arr, arr, out=out)

    def test_out_contiguous(self):
        a = np.ones((5, 2), dtype=float)
        b = np.array([[1, 3], [5, 7]], dtype=float)
        v = np.array([1, 3], dtype=float)
        tgt = np.dot(a, b)
        tgt_mv = np.dot(a, v)

        # test out non-contiguous
        out = np.ones((5, 2, 2), dtype=float)
        c = self.matmul(a, b, out=out[..., 0])
        assert c.base is out
        assert_array_equal(c, tgt)
        c = self.matmul(a, v, out=out[:, 0, 0])
        assert_array_equal(c, tgt_mv)
        c = self.matmul(v, a.T, out=out[:, 0, 0])
        assert_array_equal(c, tgt_mv)

        # test out contiguous in only last dim
        out = np.ones((10, 2), dtype=float)
        c = self.matmul(a, b, out=out[::2, :])
        assert_array_equal(c, tgt)

        # test transposes of out, args
        out = np.ones((5, 2), dtype=float)
        c = self.matmul(b.T, a.T, out=out.T)
        assert_array_equal(out, tgt)

    m1 = np.arange(15.).reshape(5, 3)
    m2 = np.arange(21.).reshape(3, 7)
    m3 = np.arange(30.).reshape(5, 6)[:, ::2]  # non-contiguous
    vc = np.arange(10.)
    vr = np.arange(6.)
    m0 = np.zeros((3, 0))

    @pytest.mark.parametrize('args', (
            # matrix-matrix
            (m1, m2), (m2.T, m1.T), (m2.T.copy(), m1.T), (m2.T, m1.T.copy()),
            # matrix-matrix-transpose, contiguous and non
            (m1, m1.T), (m1.T, m1), (m1, m3.T), (m3, m1.T),
            (m3, m3.T), (m3.T, m3),
            # matrix-matrix non-contiguous
            (m3, m2), (m2.T, m3.T), (m2.T.copy(), m3.T),
            # vector-matrix, matrix-vector, contiguous
            (m1, vr[:3]), (vc[:5], m1), (m1.T, vc[:5]), (vr[:3], m1.T),
            # vector-matrix, matrix-vector, vector non-contiguous
            (m1, vr[::2]), (vc[::2], m1), (m1.T, vc[::2]), (vr[::2], m1.T),
            # vector-matrix, matrix-vector, matrix non-contiguous
            (m3, vr[:3]), (vc[:5], m3), (m3.T, vc[:5]), (vr[:3], m3.T),
            # vector-matrix, matrix-vector, both non-contiguous
            (m3, vr[::2]), (vc[::2], m3), (m3.T, vc[::2]), (vr[::2], m3.T),
            # size == 0
            (m0, m0.T), (m0.T, m0), (m1, m0), (m0.T, m1.T),
        ))
    def test_dot_equivalent(self, args):
        r1 = np.matmul(*args)
        r2 = np.dot(*args)
        assert_equal(r1, r2)

        r3 = np.matmul(args[0].copy(), args[1].copy())
        assert_equal(r1, r3)

    def test_matmul_object(self):
        import fractions

        f = np.vectorize(fractions.Fraction)

        def random_ints():
            return np.random.randint(1, 1000, size=(10, 3, 3))
        M1 = f(random_ints(), random_ints())
        M2 = f(random_ints(), random_ints())

        M3 = self.matmul(M1, M2)

        [N1, N2, N3] = [a.astype(float) for a in [M1, M2, M3]]

        assert_allclose(N3, self.matmul(N1, N2))

    def test_matmul_object_type_scalar(self):
        from fractions import Fraction as F
        v = np.array([F(2, 3), F(5, 7)])
        res = self.matmul(v, v)
        assert_(type(res) is F)

    def test_matmul_empty(self):
        a = np.empty((3, 0), dtype=object)
        b = np.empty((0, 3), dtype=object)
        c = np.zeros((3, 3))
        assert_array_equal(np.matmul(a, b), c)

    def test_matmul_exception_multiply(self):
        # test that matmul fails if `__mul__` is missing
        class add_not_multiply:
            def __add__(self, other):
                return self
        a = np.full((3, 3), add_not_multiply())
        with assert_raises(TypeError):
            b = np.matmul(a, a)

    def test_matmul_exception_add(self):
        # test that matmul fails if `__add__` is missing
        class multiply_not_add:
            def __mul__(self, other):
                return self
        a = np.full((3, 3), multiply_not_add())
        with assert_raises(TypeError):
            b = np.matmul(a, a)

    def test_matmul_bool(self):
        # gh-14439
        a = np.array([[1, 0], [1, 1]], dtype=bool)
        assert np.max(a.view(np.uint8)) == 1
        b = np.matmul(a, a)
        # matmul with boolean output should always be 0, 1
        assert np.max(b.view(np.uint8)) == 1

        rg = np.random.default_rng(np.random.PCG64(43))
        d = rg.integers(2, size=4 * 5, dtype=np.int8)
        d = d.reshape(4, 5) > 0
        out1 = np.matmul(d, d.reshape(5, 4))
        out2 = np.dot(d, d.reshape(5, 4))
        assert_equal(out1, out2)

        c = np.matmul(np.zeros((2, 0), dtype=bool), np.zeros(0, dtype=bool))
        assert not np.any(c)


class TestMatmulOperator(MatmulCommon):
    import operator
    matmul = operator.matmul

    def test_array_priority_override(self):

        class A:
            __array_priority__ = 1000

            def __matmul__(self, other):
                return "A"

            def __rmatmul__(self, other):
                return "A"

        a = A()
        b = np.ones(2)
        assert_equal(self.matmul(a, b), "A")
        assert_equal(self.matmul(b, a), "A")

    def test_matmul_raises(self):
        assert_raises(TypeError, self.matmul, np.int8(5), np.int8(5))
        assert_raises(TypeError, self.matmul, np.void(b'abc'), np.void(b'abc'))
        assert_raises(TypeError, self.matmul, np.arange(10), np.void(b'abc'))


class TestMatmulInplace:
    DTYPES = {}
    for i in MatmulCommon.types:
        for j in MatmulCommon.types:
            if np.can_cast(j, i):
                DTYPES[f"{i}-{j}"] = (np.dtype(i), np.dtype(j))

    @pytest.mark.parametrize("dtype1,dtype2", DTYPES.values(), ids=DTYPES)
    def test_basic(self, dtype1: np.dtype, dtype2: np.dtype) -> None:
        a = np.arange(10).reshape(5, 2).astype(dtype1)
        a_id = id(a)
        b = np.ones((2, 2), dtype=dtype2)

        ref = a @ b
        a @= b

        assert id(a) == a_id
        assert a.dtype == dtype1
        assert a.shape == (5, 2)
        if dtype1.kind in "fc":
            np.testing.assert_allclose(a, ref)
        else:
            np.testing.assert_array_equal(a, ref)

    SHAPES = {
        "2d_large": ((10**5, 10), (10, 10)),
        "3d_large": ((10**4, 10, 10), (1, 10, 10)),
        "1d": ((3,), (3,)),
        "2d_1d": ((3, 3), (3,)),
        "1d_2d": ((3,), (3, 3)),
        "2d_broadcast": ((3, 3), (3, 1)),
        "2d_broadcast_reverse": ((1, 3), (3, 3)),
        "3d_broadcast1": ((3, 3, 3), (1, 3, 1)),
        "3d_broadcast2": ((3, 3, 3), (1, 3, 3)),
        "3d_broadcast3": ((3, 3, 3), (3, 3, 1)),
        "3d_broadcast_reverse1": ((1, 3, 3), (3, 3, 3)),
        "3d_broadcast_reverse2": ((3, 1, 3), (3, 3, 3)),
        "3d_broadcast_reverse3": ((1, 1, 3), (3, 3, 3)),
    }

    @pytest.mark.parametrize("a_shape,b_shape", SHAPES.values(), ids=SHAPES)
    def test_shapes(self, a_shape: tuple[int, ...], b_shape: tuple[int, ...]):
        a_size = np.prod(a_shape)
        a = np.arange(a_size).reshape(a_shape).astype(np.float64)
        a_id = id(a)

        b_size = np.prod(b_shape)
        b = np.arange(b_size).reshape(b_shape)

        ref = a @ b
        if ref.shape != a_shape:
            with pytest.raises(ValueError):
                a @= b
            return
        else:
            a @= b

        assert id(a) == a_id
        assert a.dtype.type == np.float64
        assert a.shape == a_shape
        np.testing.assert_allclose(a, ref)


def test_matmul_axes():
    a = np.arange(3 * 4 * 5).reshape(3, 4, 5)
    c = np.matmul(a, a, axes=[(-2, -1), (-1, -2), (1, 2)])
    assert c.shape == (3, 4, 4)
    d = np.matmul(a, a, axes=[(-2, -1), (-1, -2), (0, 1)])
    assert d.shape == (4, 4, 3)
    e = np.swapaxes(d, 0, 2)
    assert_array_equal(e, c)
    f = np.matmul(a, np.arange(3), axes=[(1, 0), (0), (0)])
    assert f.shape == (4, 5)


class TestInner:

    def test_inner_type_mismatch(self):
        c = 1.
        A = np.array((1, 1), dtype='i,i')

        assert_raises(TypeError, np.inner, c, A)
        assert_raises(TypeError, np.inner, A, c)

    def test_inner_scalar_and_vector(self):
        for dt in np.typecodes['AllInteger'] + np.typecodes['AllFloat'] + '?':
            sca = np.array(3, dtype=dt)[()]
            vec = np.array([1, 2], dtype=dt)
            desired = np.array([3, 6], dtype=dt)
            assert_equal(np.inner(vec, sca), desired)
            assert_equal(np.inner(sca, vec), desired)

    def test_vecself(self):
        # Ticket 844.
        # Inner product of a vector with itself segfaults or give
        # meaningless result
        a = np.zeros(shape=(1, 80), dtype=np.float64)
        p = np.inner(a, a)
        assert_almost_equal(p, 0, decimal=14)

    def test_inner_product_with_various_contiguities(self):
        # github issue 6532
        for dt in np.typecodes['AllInteger'] + np.typecodes['AllFloat'] + '?':
            # check an inner product involving a matrix transpose
            A = np.array([[1, 2], [3, 4]], dtype=dt)
            B = np.array([[1, 3], [2, 4]], dtype=dt)
            C = np.array([1, 1], dtype=dt)
            desired = np.array([4, 6], dtype=dt)
            assert_equal(np.inner(A.T, C), desired)
            assert_equal(np.inner(C, A.T), desired)
            assert_equal(np.inner(B, C), desired)
            assert_equal(np.inner(C, B), desired)
            # check a matrix product
            desired = np.array([[7, 10], [15, 22]], dtype=dt)
            assert_equal(np.inner(A, B), desired)
            # check the syrk vs. gemm paths
            desired = np.array([[5, 11], [11, 25]], dtype=dt)
            assert_equal(np.inner(A, A), desired)
            assert_equal(np.inner(A, A.copy()), desired)
            # check an inner product involving an aliased and reversed view
            a = np.arange(5).astype(dt)
            b = a[::-1]
            desired = np.array(10, dtype=dt).item()
            assert_equal(np.inner(b, a), desired)

    def test_3d_tensor(self):
        for dt in np.typecodes['AllInteger'] + np.typecodes['AllFloat'] + '?':
            a = np.arange(24).reshape(2, 3, 4).astype(dt)
            b = np.arange(24, 48).reshape(2, 3, 4).astype(dt)
            desired = np.array(
                [[[[ 158,  182,  206],
                   [ 230,  254,  278]],

                  [[ 566,  654,  742],
                   [ 830,  918, 1006]],

                  [[ 974, 1126, 1278],
                   [1430, 1582, 1734]]],

                 [[[1382, 1598, 1814],
                   [2030, 2246, 2462]],

                  [[1790, 2070, 2350],
                   [2630, 2910, 3190]],

                  [[2198, 2542, 2886],
                   [3230, 3574, 3918]]]]
            ).astype(dt)
            assert_equal(np.inner(a, b), desired)
            assert_equal(np.inner(b, a).transpose(2, 3, 0, 1), desired)


class TestChoose:
    def setup_method(self):
        self.x = 2 * np.ones((3,), dtype=int)
        self.y = 3 * np.ones((3,), dtype=int)
        self.x2 = 2 * np.ones((2, 3), dtype=int)
        self.y2 = 3 * np.ones((2, 3), dtype=int)
        self.ind = [0, 0, 1]

    def test_basic(self):
        A = np.choose(self.ind, (self.x, self.y))
        assert_equal(A, [2, 2, 3])

    def test_broadcast1(self):
        A = np.choose(self.ind, (self.x2, self.y2))
        assert_equal(A, [[2, 2, 3], [2, 2, 3]])

    def test_broadcast2(self):
        A = np.choose(self.ind, (self.x, self.y2))
        assert_equal(A, [[2, 2, 3], [2, 2, 3]])

    @pytest.mark.parametrize("ops",
        [(1000, np.array([1], dtype=np.uint8)),
         (-1, np.array([1], dtype=np.uint8)),
         (1., np.float32(3)),
         (1., np.array([3], dtype=np.float32))],)
    def test_output_dtype(self, ops):
        expected_dt = np.result_type(*ops)
        assert np.choose([0], ops).dtype == expected_dt

    def test_dimension_and_args_limit(self):
        # Maxdims for the legacy iterator is 32, but the maximum number
        # of arguments is actually larger (a itself also counts here)
        a = np.ones((1,) * 32, dtype=np.intp)
        res = a.choose([0, a] + [2] * 61)
        with pytest.raises(ValueError,
                match="Need at least 0 and at most 64 array objects"):
            a.choose([0, a] + [2] * 62)

        assert_array_equal(res, a)
        # Choose is unfortunately limited to 32 dims as of NumPy 2.0
        a = np.ones((1,) * 60, dtype=np.intp)
        with pytest.raises(RuntimeError,
                match=".*32 dimensions but the array has 60"):
            a.choose([a, a])


class TestRepeat:
    def setup_method(self):
        self.m = np.array([1, 2, 3, 4, 5, 6])
        self.m_rect = self.m.reshape((2, 3))

    def test_basic(self):
        A = np.repeat(self.m, [1, 3, 2, 1, 1, 2])
        assert_equal(A, [1, 2, 2, 2, 3,
                         3, 4, 5, 6, 6])

    def test_broadcast1(self):
        A = np.repeat(self.m, 2)
        assert_equal(A, [1, 1, 2, 2, 3, 3,
                         4, 4, 5, 5, 6, 6])

    def test_axis_spec(self):
        A = np.repeat(self.m_rect, [2, 1], axis=0)
        assert_equal(A, [[1, 2, 3],
                         [1, 2, 3],
                         [4, 5, 6]])

        A = np.repeat(self.m_rect, [1, 3, 2], axis=1)
        assert_equal(A, [[1, 2, 2, 2, 3, 3],
                         [4, 5, 5, 5, 6, 6]])

    def test_broadcast2(self):
        A = np.repeat(self.m_rect, 2, axis=0)
        assert_equal(A, [[1, 2, 3],
                         [1, 2, 3],
                         [4, 5, 6],
                         [4, 5, 6]])

        A = np.repeat(self.m_rect, 2, axis=1)
        assert_equal(A, [[1, 1, 2, 2, 3, 3],
                         [4, 4, 5, 5, 6, 6]])


# TODO: test for multidimensional
NEIGH_MODE = {'zero': 0, 'one': 1, 'constant': 2, 'circular': 3, 'mirror': 4}


@pytest.mark.parametrize('dt', [float, Decimal], ids=['float', 'object'])
class TestNeighborhoodIter:
    # Simple, 2d tests
    def test_simple2d(self, dt):
        # Test zero and one padding for simple data type
        x = np.array([[0, 1], [2, 3]], dtype=dt)
        r = [np.array([[0, 0, 0], [0, 0, 1]], dtype=dt),
             np.array([[0, 0, 0], [0, 1, 0]], dtype=dt),
             np.array([[0, 0, 1], [0, 2, 3]], dtype=dt),
             np.array([[0, 1, 0], [2, 3, 0]], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator(
                x, [-1, 0, -1, 1], x[0], NEIGH_MODE['zero'])
        assert_array_equal(l, r)

        r = [np.array([[1, 1, 1], [1, 0, 1]], dtype=dt),
             np.array([[1, 1, 1], [0, 1, 1]], dtype=dt),
             np.array([[1, 0, 1], [1, 2, 3]], dtype=dt),
             np.array([[0, 1, 1], [2, 3, 1]], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator(
                x, [-1, 0, -1, 1], x[0], NEIGH_MODE['one'])
        assert_array_equal(l, r)

        r = [np.array([[4, 4, 4], [4, 0, 1]], dtype=dt),
             np.array([[4, 4, 4], [0, 1, 4]], dtype=dt),
             np.array([[4, 0, 1], [4, 2, 3]], dtype=dt),
             np.array([[0, 1, 4], [2, 3, 4]], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator(
                x, [-1, 0, -1, 1], 4, NEIGH_MODE['constant'])
        assert_array_equal(l, r)

        # Test with start in the middle
        r = [np.array([[4, 0, 1], [4, 2, 3]], dtype=dt),
             np.array([[0, 1, 4], [2, 3, 4]], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator(
                x, [-1, 0, -1, 1], 4, NEIGH_MODE['constant'], 2)
        assert_array_equal(l, r)

    def test_mirror2d(self, dt):
        x = np.array([[0, 1], [2, 3]], dtype=dt)
        r = [np.array([[0, 0, 1], [0, 0, 1]], dtype=dt),
             np.array([[0, 1, 1], [0, 1, 1]], dtype=dt),
             np.array([[0, 0, 1], [2, 2, 3]], dtype=dt),
             np.array([[0, 1, 1], [2, 3, 3]], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator(
                x, [-1, 0, -1, 1], x[0], NEIGH_MODE['mirror'])
        assert_array_equal(l, r)

    # Simple, 1d tests
    def test_simple(self, dt):
        # Test padding with constant values
        x = np.linspace(1, 5, 5).astype(dt)
        r = [[0, 1, 2], [1, 2, 3], [2, 3, 4], [3, 4, 5], [4, 5, 0]]
        l = _multiarray_tests.test_neighborhood_iterator(
                x, [-1, 1], x[0], NEIGH_MODE['zero'])
        assert_array_equal(l, r)

        r = [[1, 1, 2], [1, 2, 3], [2, 3, 4], [3, 4, 5], [4, 5, 1]]
        l = _multiarray_tests.test_neighborhood_iterator(
                x, [-1, 1], x[0], NEIGH_MODE['one'])
        assert_array_equal(l, r)

        r = [[x[4], 1, 2], [1, 2, 3], [2, 3, 4], [3, 4, 5], [4, 5, x[4]]]
        l = _multiarray_tests.test_neighborhood_iterator(
                x, [-1, 1], x[4], NEIGH_MODE['constant'])
        assert_array_equal(l, r)

    # Test mirror modes
    def test_mirror(self, dt):
        x = np.linspace(1, 5, 5).astype(dt)
        r = np.array([[2, 1, 1, 2, 3], [1, 1, 2, 3, 4], [1, 2, 3, 4, 5],
                [2, 3, 4, 5, 5], [3, 4, 5, 5, 4]], dtype=dt)
        l = _multiarray_tests.test_neighborhood_iterator(
                x, [-2, 2], x[1], NEIGH_MODE['mirror'])
        assert_([i.dtype == dt for i in l])
        assert_array_equal(l, r)

    # Circular mode
    def test_circular(self, dt):
        x = np.linspace(1, 5, 5).astype(dt)
        r = np.array([[4, 5, 1, 2, 3], [5, 1, 2, 3, 4], [1, 2, 3, 4, 5],
                [2, 3, 4, 5, 1], [3, 4, 5, 1, 2]], dtype=dt)
        l = _multiarray_tests.test_neighborhood_iterator(
                x, [-2, 2], x[0], NEIGH_MODE['circular'])
        assert_array_equal(l, r)


# Test stacking neighborhood iterators
class TestStackedNeighborhoodIter:
    # Simple, 1d test: stacking 2 constant-padded neigh iterators
    def test_simple_const(self):
        dt = np.float64
        # Test zero and one padding for simple data type
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([0], dtype=dt),
             np.array([0], dtype=dt),
             np.array([1], dtype=dt),
             np.array([2], dtype=dt),
             np.array([3], dtype=dt),
             np.array([0], dtype=dt),
             np.array([0], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [-2, 4], NEIGH_MODE['zero'], [0, 0], NEIGH_MODE['zero'])
        assert_array_equal(l, r)

        r = [np.array([1, 0, 1], dtype=dt),
             np.array([0, 1, 2], dtype=dt),
             np.array([1, 2, 3], dtype=dt),
             np.array([2, 3, 0], dtype=dt),
             np.array([3, 0, 1], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [-1, 3], NEIGH_MODE['zero'], [-1, 1], NEIGH_MODE['one'])
        assert_array_equal(l, r)

    # 2nd simple, 1d test: stacking 2 neigh iterators, mixing const padding and
    # mirror padding
    def test_simple_mirror(self):
        dt = np.float64
        # Stacking zero on top of mirror
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([0, 1, 1], dtype=dt),
             np.array([1, 1, 2], dtype=dt),
             np.array([1, 2, 3], dtype=dt),
             np.array([2, 3, 3], dtype=dt),
             np.array([3, 3, 0], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [-1, 3], NEIGH_MODE['mirror'], [-1, 1], NEIGH_MODE['zero'])
        assert_array_equal(l, r)

        # Stacking mirror on top of zero
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([1, 0, 0], dtype=dt),
             np.array([0, 0, 1], dtype=dt),
             np.array([0, 1, 2], dtype=dt),
             np.array([1, 2, 3], dtype=dt),
             np.array([2, 3, 0], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [-1, 3], NEIGH_MODE['zero'], [-2, 0], NEIGH_MODE['mirror'])
        assert_array_equal(l, r)

        # Stacking mirror on top of zero: 2nd
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([0, 1, 2], dtype=dt),
             np.array([1, 2, 3], dtype=dt),
             np.array([2, 3, 0], dtype=dt),
             np.array([3, 0, 0], dtype=dt),
             np.array([0, 0, 3], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [-1, 3], NEIGH_MODE['zero'], [0, 2], NEIGH_MODE['mirror'])
        assert_array_equal(l, r)

        # Stacking mirror on top of zero: 3rd
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([1, 0, 0, 1, 2], dtype=dt),
             np.array([0, 0, 1, 2, 3], dtype=dt),
             np.array([0, 1, 2, 3, 0], dtype=dt),
             np.array([1, 2, 3, 0, 0], dtype=dt),
             np.array([2, 3, 0, 0, 3], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [-1, 3], NEIGH_MODE['zero'], [-2, 2], NEIGH_MODE['mirror'])
        assert_array_equal(l, r)

    # 3rd simple, 1d test: stacking 2 neigh iterators, mixing const padding and
    # circular padding
    def test_simple_circular(self):
        dt = np.float64
        # Stacking zero on top of mirror
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([0, 3, 1], dtype=dt),
             np.array([3, 1, 2], dtype=dt),
             np.array([1, 2, 3], dtype=dt),
             np.array([2, 3, 1], dtype=dt),
             np.array([3, 1, 0], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [-1, 3], NEIGH_MODE['circular'], [-1, 1], NEIGH_MODE['zero'])
        assert_array_equal(l, r)

        # Stacking mirror on top of zero
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([3, 0, 0], dtype=dt),
             np.array([0, 0, 1], dtype=dt),
             np.array([0, 1, 2], dtype=dt),
             np.array([1, 2, 3], dtype=dt),
             np.array([2, 3, 0], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [-1, 3], NEIGH_MODE['zero'], [-2, 0], NEIGH_MODE['circular'])
        assert_array_equal(l, r)

        # Stacking mirror on top of zero: 2nd
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([0, 1, 2], dtype=dt),
             np.array([1, 2, 3], dtype=dt),
             np.array([2, 3, 0], dtype=dt),
             np.array([3, 0, 0], dtype=dt),
             np.array([0, 0, 1], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [-1, 3], NEIGH_MODE['zero'], [0, 2], NEIGH_MODE['circular'])
        assert_array_equal(l, r)

        # Stacking mirror on top of zero: 3rd
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([3, 0, 0, 1, 2], dtype=dt),
             np.array([0, 0, 1, 2, 3], dtype=dt),
             np.array([0, 1, 2, 3, 0], dtype=dt),
             np.array([1, 2, 3, 0, 0], dtype=dt),
             np.array([2, 3, 0, 0, 1], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [-1, 3], NEIGH_MODE['zero'], [-2, 2], NEIGH_MODE['circular'])
        assert_array_equal(l, r)

    # 4th simple, 1d test: stacking 2 neigh iterators, but with lower iterator
    # being strictly within the array
    def test_simple_strict_within(self):
        dt = np.float64
        # Stacking zero on top of zero, first neighborhood strictly inside the
        # array
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([1, 2, 3, 0], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [1, 1], NEIGH_MODE['zero'], [-1, 2], NEIGH_MODE['zero'])
        assert_array_equal(l, r)

        # Stacking mirror on top of zero, first neighborhood strictly inside the
        # array
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([1, 2, 3, 3], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [1, 1], NEIGH_MODE['zero'], [-1, 2], NEIGH_MODE['mirror'])
        assert_array_equal(l, r)

        # Stacking mirror on top of zero, first neighborhood strictly inside the
        # array
        x = np.array([1, 2, 3], dtype=dt)
        r = [np.array([1, 2, 3, 1], dtype=dt)]
        l = _multiarray_tests.test_neighborhood_iterator_oob(
                x, [1, 1], NEIGH_MODE['zero'], [-1, 2], NEIGH_MODE['circular'])
        assert_array_equal(l, r)

class TestWarnings:

    def test_complex_warning(self):
        x = np.array([1, 2])
        y = np.array([1 - 2j, 1 + 2j])

        with warnings.catch_warnings():
            warnings.simplefilter("error", ComplexWarning)
            assert_raises(ComplexWarning, x.__setitem__, slice(None), y)
            assert_equal(x, [1, 2])


class TestMinScalarType:

    def test_usigned_shortshort(self):
        dt = np.min_scalar_type(2**8 - 1)
        wanted = np.dtype('uint8')
        assert_equal(wanted, dt)

    def test_usigned_short(self):
        dt = np.min_scalar_type(2**16 - 1)
        wanted = np.dtype('uint16')
        assert_equal(wanted, dt)

    def test_usigned_int(self):
        dt = np.min_scalar_type(2**32 - 1)
        wanted = np.dtype('uint32')
        assert_equal(wanted, dt)

    def test_usigned_longlong(self):
        dt = np.min_scalar_type(2**63 - 1)
        wanted = np.dtype('uint64')
        assert_equal(wanted, dt)

    def test_object(self):
        dt = np.min_scalar_type(2**64)
        wanted = np.dtype('O')
        assert_equal(wanted, dt)


from numpy._core._internal import _dtype_from_pep3118


class TestPEP3118Dtype:
    def _check(self, spec, wanted):
        dt = np.dtype(wanted)
        actual = _dtype_from_pep3118(spec)
        assert_equal(actual, dt,
                     err_msg=f"spec {spec!r} != dtype {wanted!r}")

    def test_native_padding(self):
        align = np.dtype('i').alignment
        for j in range(8):
            if j == 0:
                s = 'bi'
            else:
                s = 'b%dxi' % j
            self._check('@' + s, {'f0': ('i1', 0),
                                'f1': ('i', align * (1 + j // align))})
            self._check('=' + s, {'f0': ('i1', 0),
                                'f1': ('i', 1 + j)})

    def test_native_padding_2(self):
        # Native padding should work also for structs and sub-arrays
        self._check('x3T{xi}', {'f0': (({'f0': ('i', 4)}, (3,)), 4)})
        self._check('^x3T{xi}', {'f0': (({'f0': ('i', 1)}, (3,)), 1)})

    def test_trailing_padding(self):
        # Trailing padding should be included, *and*, the item size
        # should match the alignment if in aligned mode
        align = np.dtype('i').alignment
        size = np.dtype('i').itemsize

        def aligned(n):
            return align * (1 + (n - 1) // align)

        base = {"formats": ['i'], "names": ['f0']}

        self._check('ix',    dict(itemsize=aligned(size + 1), **base))
        self._check('ixx',   dict(itemsize=aligned(size + 2), **base))
        self._check('ixxx',  dict(itemsize=aligned(size + 3), **base))
        self._check('ixxxx', dict(itemsize=aligned(size + 4), **base))
        self._check('i7x',   dict(itemsize=aligned(size + 7), **base))

        self._check('^ix',    dict(itemsize=size + 1, **base))
        self._check('^ixx',   dict(itemsize=size + 2, **base))
        self._check('^ixxx',  dict(itemsize=size + 3, **base))
        self._check('^ixxxx', dict(itemsize=size + 4, **base))
        self._check('^i7x',   dict(itemsize=size + 7, **base))

    def test_native_padding_3(self):
        dt = np.dtype(
                [('a', 'b'), ('b', 'i'),
                    ('sub', np.dtype('b,i')), ('c', 'i')],
                align=True)
        self._check("T{b:a:xxxi:b:T{b:f0:=i:f1:}:sub:xxxi:c:}", dt)

        dt = np.dtype(
                [('a', 'b'), ('b', 'i'), ('c', 'b'), ('d', 'b'),
                    ('e', 'b'), ('sub', np.dtype('b,i', align=True))])
        self._check("T{b:a:=i:b:b:c:b:d:b:e:T{b:f0:xxxi:f1:}:sub:}", dt)

    def test_padding_with_array_inside_struct(self):
        dt = np.dtype(
                [('a', 'b'), ('b', 'i'), ('c', 'b', (3,)),
                    ('d', 'i')],
                align=True)
        self._check("T{b:a:xxxi:b:3b:c:xi:d:}", dt)

    def test_byteorder_inside_struct(self):
        # The byte order after @T{=i} should be '=', not '@'.
        # Check this by noting the absence of native alignment.
        self._check('@T{^i}xi', {'f0': ({'f0': ('i', 0)}, 0),
                                 'f1': ('i', 5)})

    def test_intra_padding(self):
        # Natively aligned sub-arrays may require some internal padding
        align = np.dtype('i').alignment
        size = np.dtype('i').itemsize

        def aligned(n):
            return (align * (1 + (n - 1) // align))

        self._check('(3)T{ix}', ({
            "names": ['f0'],
            "formats": ['i'],
            "offsets": [0],
            "itemsize": aligned(size + 1)
        }, (3,)))

    def test_char_vs_string(self):
        dt = np.dtype('c')
        self._check('c', dt)

        dt = np.dtype([('f0', 'S1', (4,)), ('f1', 'S4')])
        self._check('4c4s', dt)

    def test_field_order(self):
        # gh-9053 - previously, we relied on dictionary key order
        self._check("(0)I:a:f:b:", [('a', 'I', (0,)), ('b', 'f')])
        self._check("(0)I:b:f:a:", [('b', 'I', (0,)), ('a', 'f')])

    def test_unnamed_fields(self):
        self._check('ii',     [('f0', 'i'), ('f1', 'i')])
        self._check('ii:f0:', [('f1', 'i'), ('f0', 'i')])

        self._check('i', 'i')
        self._check('i:f0:', [('f0', 'i')])


class TestNewBufferProtocol:
    """ Test PEP3118 buffers """

    def _check_roundtrip(self, obj):
        obj = np.asarray(obj)
        x = memoryview(obj)
        y = np.asarray(x)
        y2 = np.array(x)
        assert_(not y.flags.owndata)
        assert_(y2.flags.owndata)

        assert_equal(y.dtype, obj.dtype)
        assert_equal(y.shape, obj.shape)
        assert_array_equal(obj, y)

        assert_equal(y2.dtype, obj.dtype)
        assert_equal(y2.shape, obj.shape)
        assert_array_equal(obj, y2)

    def test_roundtrip(self):
        x = np.array([1, 2, 3, 4, 5], dtype='i4')
        self._check_roundtrip(x)

        x = np.array([[1, 2], [3, 4]], dtype=np.float64)
        self._check_roundtrip(x)

        x = np.zeros((3, 3, 3), dtype=np.float32)[:, 0, :]
        self._check_roundtrip(x)

        dt = [('a', 'b'),
              ('b', 'h'),
              ('c', 'i'),
              ('d', 'l'),
              ('dx', 'q'),
              ('e', 'B'),
              ('f', 'H'),
              ('g', 'I'),
              ('h', 'L'),
              ('hx', 'Q'),
              ('i', np.single),
              ('j', np.double),
              ('k', np.longdouble),
              ('ix', np.csingle),
              ('jx', np.cdouble),
              ('kx', np.clongdouble),
              ('l', 'S4'),
              ('m', 'U4'),
              ('n', 'V3'),
              ('o', '?'),
              ('p', np.half),
              ]
        x = np.array(
                [(1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                    b'aaaa', 'bbbb', b'xxx', True, 1.0)],
                dtype=dt)
        self._check_roundtrip(x)

        x = np.array(([[1, 2], [3, 4]],), dtype=[('a', (int, (2, 2)))])
        self._check_roundtrip(x)

        x = np.array([1, 2, 3], dtype='>i2')
        self._check_roundtrip(x)

        x = np.array([1, 2, 3], dtype='<i2')
        self._check_roundtrip(x)

        x = np.array([1, 2, 3], dtype='>i4')
        self._check_roundtrip(x)

        x = np.array([1, 2, 3], dtype='<i4')
        self._check_roundtrip(x)

        # check long long can be represented as non-native
        x = np.array([1, 2, 3], dtype='>q')
        self._check_roundtrip(x)

        # Native-only data types can be passed through the buffer interface
        # only in native byte order
        if sys.byteorder == 'little':
            x = np.array([1, 2, 3], dtype='>g')
            assert_raises(ValueError, self._check_roundtrip, x)
            x = np.array([1, 2, 3], dtype='<g')
            self._check_roundtrip(x)
        else:
            x = np.array([1, 2, 3], dtype='>g')
            self._check_roundtrip(x)
            x = np.array([1, 2, 3], dtype='<g')
            assert_raises(ValueError, self._check_roundtrip, x)

    def test_roundtrip_half(self):
        half_list = [
            1.0,
            -2.0,
            6.5504 * 10**4,  # (max half precision)
            2**-14,  # ~= 6.10352 * 10**-5 (minimum positive normal)
            2**-24,  # ~= 5.96046 * 10**-8 (minimum strictly positive subnormal)
            0.0,
            -0.0,
            float('+inf'),
            float('-inf'),
            0.333251953125,  # ~= 1/3
        ]

        x = np.array(half_list, dtype='>e')
        self._check_roundtrip(x)
        x = np.array(half_list, dtype='<e')
        self._check_roundtrip(x)

    def test_roundtrip_single_types(self):
        for typ in np._core.sctypeDict.values():
            dtype = np.dtype(typ)

            if dtype.char in 'Mm':
                # datetimes cannot be used in buffers
                continue
            if dtype.char == 'V':
                # skip void
                continue

            x = np.zeros(4, dtype=dtype)
            self._check_roundtrip(x)

            if dtype.char not in 'qQgG':
                dt = dtype.newbyteorder('<')
                x = np.zeros(4, dtype=dt)
                self._check_roundtrip(x)

                dt = dtype.newbyteorder('>')
                x = np.zeros(4, dtype=dt)
                self._check_roundtrip(x)

    def test_roundtrip_scalar(self):
        # Issue #4015.
        self._check_roundtrip(0)

    def test_invalid_buffer_format(self):
        # datetime64 cannot be used fully in a buffer yet
        # Should be fixed in the next Numpy major release
        dt = np.dtype([('a', 'uint16'), ('b', 'M8[s]')])
        a = np.empty(3, dt)
        assert_raises((ValueError, BufferError), memoryview, a)
        assert_raises((ValueError, BufferError), memoryview, np.array((3), 'M8[D]'))

    def test_export_simple_1d(self):
        x = np.array([1, 2, 3, 4, 5], dtype='i')
        y = memoryview(x)
        assert_equal(y.format, 'i')
        assert_equal(y.shape, (5,))
        assert_equal(y.ndim, 1)
        assert_equal(y.strides, (4,))
        assert_equal(y.suboffsets, ())
        assert_equal(y.itemsize, 4)

    def test_export_simple_nd(self):
        x = np.array([[1, 2], [3, 4]], dtype=np.float64)
        y = memoryview(x)
        assert_equal(y.format, 'd')
        assert_equal(y.shape, (2, 2))
        assert_equal(y.ndim, 2)
        assert_equal(y.strides, (16, 8))
        assert_equal(y.suboffsets, ())
        assert_equal(y.itemsize, 8)

    def test_export_discontiguous(self):
        x = np.zeros((3, 3, 3), dtype=np.float32)[:, 0, :]
        y = memoryview(x)
        assert_equal(y.format, 'f')
        assert_equal(y.shape, (3, 3))
        assert_equal(y.ndim, 2)
        assert_equal(y.strides, (36, 4))
        assert_equal(y.suboffsets, ())
        assert_equal(y.itemsize, 4)

    def test_export_record(self):
        dt = [('a', 'b'),
              ('b', 'h'),
              ('c', 'i'),
              ('d', 'l'),
              ('dx', 'q'),
              ('e', 'B'),
              ('f', 'H'),
              ('g', 'I'),
              ('h', 'L'),
              ('hx', 'Q'),
              ('i', np.single),
              ('j', np.double),
              ('k', np.longdouble),
              ('ix', np.csingle),
              ('jx', np.cdouble),
              ('kx', np.clongdouble),
              ('l', 'S4'),
              ('m', 'U4'),
              ('n', 'V3'),
              ('o', '?'),
              ('p', np.half),
              ]
        x = np.array(
                [(1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                    b'aaaa', 'bbbb', b'   ', True, 1.0)],
                dtype=dt)
        y = memoryview(x)
        assert_equal(y.shape, (1,))
        assert_equal(y.ndim, 1)
        assert_equal(y.suboffsets, ())

        sz = sum(np.dtype(b).itemsize for a, b in dt)
        if np.dtype('l').itemsize == 4:
            assert_equal(y.format, 'T{b:a:=h:b:i:c:l:d:q:dx:B:e:@H:f:=I:g:L:h:Q:hx:f:i:d:j:^g:k:=Zf:ix:Zd:jx:^Zg:kx:4s:l:=4w:m:3x:n:?:o:@e:p:}')
        else:
            assert_equal(y.format, 'T{b:a:=h:b:i:c:q:d:q:dx:B:e:@H:f:=I:g:Q:h:Q:hx:f:i:d:j:^g:k:=Zf:ix:Zd:jx:^Zg:kx:4s:l:=4w:m:3x:n:?:o:@e:p:}')
        assert_equal(y.strides, (sz,))
        assert_equal(y.itemsize, sz)

    def test_export_subarray(self):
        x = np.array(([[1, 2], [3, 4]],), dtype=[('a', ('i', (2, 2)))])
        y = memoryview(x)
        assert_equal(y.format, 'T{(2,2)i:a:}')
        assert_equal(y.shape, ())
        assert_equal(y.ndim, 0)
        assert_equal(y.strides, ())
        assert_equal(y.suboffsets, ())
        assert_equal(y.itemsize, 16)

    def test_export_endian(self):
        x = np.array([1, 2, 3], dtype='>i')
        y = memoryview(x)
        if sys.byteorder == 'little':
            assert_equal(y.format, '>i')
        else:
            assert_equal(y.format, 'i')

        x = np.array([1, 2, 3], dtype='<i')
        y = memoryview(x)
        if sys.byteorder == 'little':
            assert_equal(y.format, 'i')
        else:
            assert_equal(y.format, '<i')

    def test_export_flags(self):
        # Check SIMPLE flag, see also gh-3613 (exception should be BufferError)
        assert_raises(ValueError,
                      _multiarray_tests.get_buffer_info,
                       np.arange(5)[::2], ('SIMPLE',))

    @pytest.mark.parametrize(["obj", "error"], [
            pytest.param(np.array([1, 2], dtype=rational), ValueError, id="array"),
            pytest.param(rational(1, 2), TypeError, id="scalar")])
    def test_export_and_pickle_user_dtype(self, obj, error):
        # User dtypes should export successfully when FORMAT was not requested.
        with pytest.raises(error):
            _multiarray_tests.get_buffer_info(obj, ("STRIDED_RO", "FORMAT"))

        _multiarray_tests.get_buffer_info(obj, ("STRIDED_RO",))

        # This is currently also necessary to implement pickling:
        pickle_obj = pickle.dumps(obj)
        res = pickle.loads(pickle_obj)
        assert_array_equal(res, obj)

    def test_repr_user_dtype(self):
        dt = np.dtype(rational)
        assert_equal(repr(dt), 'dtype(rational)')

    def test_padding(self):
        for j in range(8):
            x = np.array([(1,), (2,)], dtype={'f0': (int, j)})
            self._check_roundtrip(x)

    def test_reference_leak(self):
        if HAS_REFCOUNT:
            count_1 = sys.getrefcount(np._core._internal)
        a = np.zeros(4)
        b = memoryview(a)
        c = np.asarray(b)
        if HAS_REFCOUNT:
            count_2 = sys.getrefcount(np._core._internal)
            assert_equal(count_1, count_2)

    def test_padded_struct_array(self):
        dt1 = np.dtype(
                [('a', 'b'), ('b', 'i'), ('sub', np.dtype('b,i')), ('c', 'i')],
                align=True)
        x1 = np.arange(dt1.itemsize, dtype=np.int8).view(dt1)
        self._check_roundtrip(x1)

        dt2 = np.dtype(
                [('a', 'b'), ('b', 'i'), ('c', 'b', (3,)), ('d', 'i')],
                align=True)
        x2 = np.arange(dt2.itemsize, dtype=np.int8).view(dt2)
        self._check_roundtrip(x2)

        dt3 = np.dtype(
                [('a', 'b'), ('b', 'i'), ('c', 'b'), ('d', 'b'),
                    ('e', 'b'), ('sub', np.dtype('b,i', align=True))])
        x3 = np.arange(dt3.itemsize, dtype=np.int8).view(dt3)
        self._check_roundtrip(x3)

    @pytest.mark.valgrind_error(reason="leaks buffer info cache temporarily.")
    def test_relaxed_strides(self, c=np.ones((1, 10, 10), dtype='i8')):  # noqa: B008
        # Note: c defined as parameter so that it is persistent and leak
        # checks will notice gh-16934 (buffer info cache leak).
        c.strides = (-1, 80, 8)  # strides need to be fixed at export

        assert_(memoryview(c).strides == (800, 80, 8))

        # Writing C-contiguous data to a BytesIO buffer should work
        fd = io.BytesIO()
        fd.write(c.data)

        fortran = c.T
        assert_(memoryview(fortran).strides == (8, 80, 800))

        arr = np.ones((1, 10))
        if arr.flags.f_contiguous:
            shape, strides = _multiarray_tests.get_buffer_info(
                    arr, ['F_CONTIGUOUS'])
            assert_(strides[0] == 8)
            arr = np.ones((10, 1), order='F')
            shape, strides = _multiarray_tests.get_buffer_info(
                    arr, ['C_CONTIGUOUS'])
            assert_(strides[-1] == 8)

    def test_out_of_order_fields(self):
        dt = np.dtype({
            "formats": ['<i4', '<i4'],
            "names": ['one', 'two'],
            "offsets": [4, 0],
            "itemsize": 8
        })

        # overlapping fields cannot be represented by PEP3118
        arr = np.empty(1, dt)
        with assert_raises(ValueError):
            memoryview(arr)

    def test_max_dims(self):
        a = np.ones((1,) * 32)
        self._check_roundtrip(a)

    def test_error_pointer_type(self):
        # gh-6741
        m = memoryview(ctypes.pointer(ctypes.c_uint8()))
        assert_('&' in m.format)

        assert_raises_regex(
            ValueError, "format string",
            np.array, m)

    def test_error_message_unsupported(self):
        # wchar has no corresponding numpy type - if this changes in future, we
        # need a better way to construct an invalid memoryview format.
        t = ctypes.c_wchar * 4
        with assert_raises(ValueError) as cm:
            np.array(t())

        exc = cm.exception
        with assert_raises_regex(
            NotImplementedError,
            r"Unrepresentable .* 'u' \(UCS-2 strings\)"
        ):
            raise exc.__cause__

    def test_ctypes_integer_via_memoryview(self):
        # gh-11150, due to bpo-10746
        for c_integer in {ctypes.c_int, ctypes.c_long, ctypes.c_longlong}:
            value = c_integer(42)
            with warnings.catch_warnings(record=True):
                warnings.filterwarnings('always', r'.*\bctypes\b', RuntimeWarning)
                np.asarray(value)

    def test_ctypes_struct_via_memoryview(self):
        # gh-10528
        class foo(ctypes.Structure):
            _fields_ = [('a', ctypes.c_uint8), ('b', ctypes.c_uint32)]
        f = foo(a=1, b=2)

        with warnings.catch_warnings(record=True):
            warnings.filterwarnings('always', r'.*\bctypes\b', RuntimeWarning)
            arr = np.asarray(f)

        assert_equal(arr['a'], 1)
        assert_equal(arr['b'], 2)
        f.a = 3
        assert_equal(arr['a'], 3)

    @pytest.mark.parametrize("obj", [np.ones(3), np.ones(1, dtype="i,i")[()]])
    def test_error_if_stored_buffer_info_is_corrupted(self, obj):
        """
        If a user extends a NumPy array before 1.20 and then runs it
        on NumPy 1.20+. A C-subclassed array might in theory modify
        the new buffer-info field. This checks that an error is raised
        if this happens (for buffer export), an error is written on delete.
        This is a sanity check to help users transition to safe code, it
        may be deleted at any point.
        """
        # corrupt buffer info:
        _multiarray_tests.corrupt_or_fix_bufferinfo(obj)
        name = type(obj)
        with pytest.raises(RuntimeError,
                    match=f".*{name} appears to be C subclassed"):
            memoryview(obj)
        # Fix buffer info again before we delete (or we lose the memory)
        _multiarray_tests.corrupt_or_fix_bufferinfo(obj)

    def test_no_suboffsets(self):
        try:
            import _testbuffer
        except ImportError:
            raise pytest.skip("_testbuffer is not available")

        for shape in [(2, 3), (2, 3, 4)]:
            data = list(range(np.prod(shape)))
            buffer = _testbuffer.ndarray(data, shape, format='i',
                                         flags=_testbuffer.ND_PIL)
            msg = "NumPy currently does not support.*suboffsets"
            with pytest.raises(BufferError, match=msg):
                np.asarray(buffer)
            with pytest.raises(BufferError, match=msg):
                np.asarray([buffer])

            # Also check (unrelated and more limited but similar) frombuffer:
            with pytest.raises(BufferError):
                np.frombuffer(buffer)


class TestArrayCreationCopyArgument:

    class RaiseOnBool:

        def __bool__(self):
            raise ValueError

    true_vals = [True, np._CopyMode.ALWAYS, np.True_]
    if_needed_vals = [None, np._CopyMode.IF_NEEDED]
    false_vals = [False, np._CopyMode.NEVER, np.False_]

    def test_scalars(self):
        # Test both numpy and python scalars
        for dtype in np.typecodes["All"]:
            arr = np.zeros((), dtype=dtype)
            scalar = arr[()]
            pyscalar = arr.item(0)

            # Test never-copy raises error:
            assert_raises(ValueError, np.array, pyscalar,
                            copy=self.RaiseOnBool())
            assert_raises(ValueError, _multiarray_tests.npy_ensurenocopy,
                            [1])
            for copy in self.false_vals:
                assert_raises(ValueError, np.array, scalar, copy=copy)
                assert_raises(ValueError, np.array, pyscalar, copy=copy)
                # Casting with a dtype (to unsigned integers) can be special:
                with pytest.raises(ValueError):
                    np.array(pyscalar, dtype=np.int64, copy=copy)

    def test_compatible_cast(self):

        # Some types are compatible even though they are different, no
        # copy is necessary for them. This is mostly true for some integers
        def int_types(byteswap=False):
            int_types = (np.typecodes["Integer"] +
                         np.typecodes["UnsignedInteger"])
            for int_type in int_types:
                yield np.dtype(int_type)
                if byteswap:
                    yield np.dtype(int_type).newbyteorder()

        for int1 in int_types():
            for int2 in int_types(True):
                arr = np.arange(10, dtype=int1)

                for copy in self.true_vals:
                    res = np.array(arr, copy=copy, dtype=int2)
                    assert res is not arr and res.flags.owndata
                    assert_array_equal(res, arr)

                if int1 == int2:
                    # Casting is not necessary, base check is sufficient here
                    for copy in self.if_needed_vals:
                        res = np.array(arr, copy=copy, dtype=int2)
                        assert res is arr or res.base is arr

                    for copy in self.false_vals:
                        res = np.array(arr, copy=copy, dtype=int2)
                        assert res is arr or res.base is arr

                else:
                    # Casting is necessary, assert copy works:
                    for copy in self.if_needed_vals:
                        res = np.array(arr, copy=copy, dtype=int2)
                        assert res is not arr and res.flags.owndata
                        assert_array_equal(res, arr)

                    assert_raises(ValueError, np.array,
                                  arr, copy=False,
                                  dtype=int2)

    def test_buffer_interface(self):

        # Buffer interface gives direct memory access (no copy)
        arr = np.arange(10)
        view = memoryview(arr)

        # Checking bases is a bit tricky since numpy creates another
        # memoryview, so use may_share_memory.
        for copy in self.true_vals:
            res = np.array(view, copy=copy)
            assert not np.may_share_memory(arr, res)
        for copy in self.false_vals:
            res = np.array(view, copy=copy)
            assert np.may_share_memory(arr, res)
        res = np.array(view, copy=np._CopyMode.NEVER)
        assert np.may_share_memory(arr, res)

    def test_array_interfaces(self):
        base_arr = np.arange(10)

        # Array interface gives direct memory access (much like a memoryview)
        class ArrayLike:
            __array_interface__ = base_arr.__array_interface__

        arr = ArrayLike()

        for copy, val in [(True, None), (np._CopyMode.ALWAYS, None),
                          (False, arr), (np._CopyMode.IF_NEEDED, arr),
                          (np._CopyMode.NEVER, arr)]:
            res = np.array(arr, copy=copy)
            assert res.base is val

    def test___array__(self):
        base_arr = np.arange(10)

        class ArrayLike:
            def __array__(self, dtype=None, copy=None):
                return base_arr

        arr = ArrayLike()

        for copy in self.true_vals:
            res = np.array(arr, copy=copy)
            assert_array_equal(res, base_arr)
            # An additional copy is no longer forced by NumPy in this case.
            # NumPy trusts the ArrayLike made a copy:
            assert res is base_arr

        for copy in self.if_needed_vals + self.false_vals:
            res = np.array(arr, copy=copy)
            assert_array_equal(res, base_arr)
            assert res is base_arr  # numpy trusts the ArrayLike

    def test___array__copy_arg(self):
        a = np.ones((10, 10), dtype=int)

        assert np.shares_memory(a, a.__array__())
        assert not np.shares_memory(a, a.__array__(float))
        assert not np.shares_memory(a, a.__array__(float, copy=None))
        assert not np.shares_memory(a, a.__array__(copy=True))
        assert np.shares_memory(a, a.__array__(copy=None))
        assert np.shares_memory(a, a.__array__(copy=False))
        assert np.shares_memory(a, a.__array__(int, copy=False))
        with pytest.raises(ValueError):
            np.shares_memory(a, a.__array__(float, copy=False))

        base_arr = np.arange(10)

        class ArrayLikeNoCopy:
            def __array__(self, dtype=None):
                return base_arr

        a = ArrayLikeNoCopy()

        # explicitly passing copy=None shouldn't raise a warning
        arr = np.array(a, copy=None)
        assert_array_equal(arr, base_arr)
        assert arr is base_arr

        # As of NumPy 2.1, explicitly passing copy=True does trigger passing
        # it to __array__ (deprecation warning is triggered).
        with pytest.warns(DeprecationWarning,
                          match="__array__.*must implement.*'copy'"):
            arr = np.array(a, copy=True)
        assert_array_equal(arr, base_arr)
        assert arr is not base_arr

        # And passing copy=False gives a deprecation warning, but also raises
        # an error:
        with pytest.warns(DeprecationWarning, match="__array__.*'copy'"):
            with pytest.raises(ValueError,
                    match=r"Unable to avoid copy(.|\n)*numpy_2_0_migration_guide.html"):
                np.array(a, copy=False)

    def test___array__copy_once(self):
        size = 100
        base_arr = np.zeros((size, size))
        copy_arr = np.zeros((size, size))

        class ArrayRandom:
            def __init__(self):
                self.true_passed = False

            def __array__(self, dtype=None, copy=None):
                if copy:
                    self.true_passed = True
                    return copy_arr
                else:
                    return base_arr

        arr_random = ArrayRandom()
        first_copy = np.array(arr_random, copy=True)
        assert arr_random.true_passed
        assert first_copy is copy_arr

        arr_random = ArrayRandom()
        no_copy = np.array(arr_random, copy=False)
        assert not arr_random.true_passed
        assert no_copy is base_arr

        arr_random = ArrayRandom()
        _ = np.array([arr_random], copy=True)
        assert not arr_random.true_passed

        arr_random = ArrayRandom()
        second_copy = np.array(arr_random, copy=True, order="F")
        assert arr_random.true_passed
        assert second_copy is not copy_arr

        arr_random = ArrayRandom()
        arr = np.ones((size, size))
        arr[...] = arr_random
        assert not arr_random.true_passed
        assert not np.shares_memory(arr, base_arr)

    @pytest.mark.skipif(not HAS_REFCOUNT, reason="Python lacks refcounts")
    def test__array__reference_leak(self):
        class NotAnArray:
            def __array__(self, dtype=None, copy=None):
                raise NotImplementedError

        x = NotAnArray()

        refcount = sys.getrefcount(x)

        try:
            np.array(x)
        except NotImplementedError:
            pass

        gc.collect()

        assert refcount == sys.getrefcount(x)

    @pytest.mark.parametrize(
            "arr", [np.ones(()), np.arange(81).reshape((9, 9))])
    @pytest.mark.parametrize("order1", ["C", "F", None])
    @pytest.mark.parametrize("order2", ["C", "F", "A", "K"])
    def test_order_mismatch(self, arr, order1, order2):
        # The order is the main (python side) reason that can cause
        # a never-copy to fail.
        # Prepare C-order, F-order and non-contiguous arrays:
        arr = arr.copy(order1)
        if order1 == "C":
            assert arr.flags.c_contiguous
        elif order1 == "F":
            assert arr.flags.f_contiguous
        elif arr.ndim != 0:
            # Make array non-contiguous
            arr = arr[::2, ::2]
            assert not arr.flags.forc

        # Whether a copy is necessary depends on the order of arr:
        if order2 == "C":
            no_copy_necessary = arr.flags.c_contiguous
        elif order2 == "F":
            no_copy_necessary = arr.flags.f_contiguous
        else:
            # Keeporder and Anyorder are OK with non-contiguous output.
            # This is not consistent with the `astype` behaviour which
            # enforces contiguity for "A". It is probably historic from when
            # "K" did not exist.
            no_copy_necessary = True

        # Test it for both the array and a memoryview
        for view in [arr, memoryview(arr)]:
            for copy in self.true_vals:
                res = np.array(view, copy=copy, order=order2)
                assert res is not arr and res.flags.owndata
                assert_array_equal(arr, res)

            if no_copy_necessary:
                for copy in self.if_needed_vals + self.false_vals:
                    res = np.array(view, copy=copy, order=order2)
                    # res.base.obj refers to the memoryview
                    if not IS_PYPY:
                        assert res is arr or res.base.obj is arr
            else:
                for copy in self.if_needed_vals:
                    res = np.array(arr, copy=copy, order=order2)
                    assert_array_equal(arr, res)
                for copy in self.false_vals:
                    assert_raises(ValueError, np.array,
                                  view, copy=copy, order=order2)

    def test_striding_not_ok(self):
        arr = np.array([[1, 2, 4], [3, 4, 5]])
        assert_raises(ValueError, np.array,
                      arr.T, copy=np._CopyMode.NEVER,
                      order='C')
        assert_raises(ValueError, np.array,
                      arr.T, copy=np._CopyMode.NEVER,
                      order='C', dtype=np.int64)
        assert_raises(ValueError, np.array,
                      arr, copy=np._CopyMode.NEVER,
                      order='F')
        assert_raises(ValueError, np.array,
                      arr, copy=np._CopyMode.NEVER,
                      order='F', dtype=np.int64)


class TestArrayAttributeDeletion:

    def test_multiarray_writable_attributes_deletion(self):
        # ticket #2046, should not seqfault, raise AttributeError
        a = np.ones(2)
        attr = ['shape', 'strides', 'data', 'dtype', 'real', 'imag', 'flat']
        with suppress_warnings() as sup:
            sup.filter(DeprecationWarning, "Assigning the 'data' attribute")
            for s in attr:
                assert_raises(AttributeError, delattr, a, s)

    def test_multiarray_not_writable_attributes_deletion(self):
        a = np.ones(2)
        attr = ["ndim", "flags", "itemsize", "size", "nbytes", "base",
                "ctypes", "T", "__array_interface__", "__array_struct__",
                "__array_priority__", "__array_finalize__"]
        for s in attr:
            assert_raises(AttributeError, delattr, a, s)

    def test_multiarray_flags_writable_attribute_deletion(self):
        a = np.ones(2).flags
        attr = ['writebackifcopy', 'updateifcopy', 'aligned', 'writeable']
        for s in attr:
            assert_raises(AttributeError, delattr, a, s)

    def test_multiarray_flags_not_writable_attribute_deletion(self):
        a = np.ones(2).flags
        attr = ["contiguous", "c_contiguous", "f_contiguous", "fortran",
                "owndata", "fnc", "forc", "behaved", "carray", "farray",
                "num"]
        for s in attr:
            assert_raises(AttributeError, delattr, a, s)


class TestArrayInterface:
    class Foo:
        def __init__(self, value):
            self.value = value
            self.iface = {'typestr': 'f8'}

        def __float__(self):
            return float(self.value)

        @property
        def __array_interface__(self):
            return self.iface

    f = Foo(0.5)

    @pytest.mark.parametrize('val, iface, expected', [
        (f, {}, 0.5),
        ([f], {}, [0.5]),
        ([f, f], {}, [0.5, 0.5]),
        (f, {'shape': ()}, 0.5),
        (f, {'shape': None}, TypeError),
        (f, {'shape': (1, 1)}, [[0.5]]),
        (f, {'shape': (2,)}, ValueError),
        (f, {'strides': ()}, 0.5),
        (f, {'strides': (2,)}, ValueError),
        (f, {'strides': 16}, TypeError),
        ])
    def test_scalar_interface(self, val, iface, expected):
        # Test scalar coercion within the array interface
        self.f.iface = {'typestr': 'f8'}
        self.f.iface.update(iface)
        if HAS_REFCOUNT:
            pre_cnt = sys.getrefcount(np.dtype('f8'))
        if isinstance(expected, type):
            assert_raises(expected, np.array, val)
        else:
            result = np.array(val)
            assert_equal(np.array(val), expected)
            assert result.dtype == 'f8'
            del result
        if HAS_REFCOUNT:
            post_cnt = sys.getrefcount(np.dtype('f8'))
            assert_equal(pre_cnt, post_cnt)

def test_interface_no_shape():
    class ArrayLike:
        array = np.array(1)
        __array_interface__ = array.__array_interface__
    assert_equal(np.array(ArrayLike()), 1)


def test_array_interface_itemsize():
    # See gh-6361
    my_dtype = np.dtype({'names': ['A', 'B'], 'formats': ['f4', 'f4'],
                         'offsets': [0, 8], 'itemsize': 16})
    a = np.ones(10, dtype=my_dtype)
    descr_t = np.dtype(a.__array_interface__['descr'])
    typestr_t = np.dtype(a.__array_interface__['typestr'])
    assert_equal(descr_t.itemsize, typestr_t.itemsize)


def test_array_interface_empty_shape():
    # See gh-7994
    arr = np.array([1, 2, 3])
    interface1 = dict(arr.__array_interface__)
    interface1['shape'] = ()

    class DummyArray1:
        __array_interface__ = interface1

    # NOTE: Because Py2 str/Py3 bytes supports the buffer interface, setting
    # the interface data to bytes would invoke the bug this tests for, that
    # __array_interface__ with shape=() is not allowed if the data is an object
    # exposing the buffer interface
    interface2 = dict(interface1)
    interface2['data'] = arr[0].tobytes()

    class DummyArray2:
        __array_interface__ = interface2

    arr1 = np.asarray(DummyArray1())
    arr2 = np.asarray(DummyArray2())
    arr3 = arr[:1].reshape(())
    assert_equal(arr1, arr2)
    assert_equal(arr1, arr3)

def test_array_interface_offset():
    arr = np.array([1, 2, 3], dtype='int32')
    interface = dict(arr.__array_interface__)
    interface['data'] = memoryview(arr)
    interface['shape'] = (2,)
    interface['offset'] = 4

    class DummyArray:
        __array_interface__ = interface

    arr1 = np.asarray(DummyArray())
    assert_equal(arr1, arr[1:])

def test_array_interface_unicode_typestr():
    arr = np.array([1, 2, 3], dtype='int32')
    interface = dict(arr.__array_interface__)
    interface['typestr'] = '\N{check mark}'

    class DummyArray:
        __array_interface__ = interface

    # should not be UnicodeEncodeError
    with pytest.raises(TypeError):
        np.asarray(DummyArray())

def test_flat_element_deletion():
    it = np.ones(3).flat
    try:
        del it[1]
        del it[1:2]
    except TypeError:
        pass
    except Exception:
        raise AssertionError


def test_scalar_element_deletion():
    a = np.zeros(2, dtype=[('x', 'int'), ('y', 'int')])
    assert_raises(ValueError, a[0].__delitem__, 'x')


class TestAsCArray:
    def test_1darray(self):
        array = np.arange(24, dtype=np.double)
        from_c = _multiarray_tests.test_as_c_array(array, 3)
        assert_equal(array[3], from_c)

    def test_2darray(self):
        array = np.arange(24, dtype=np.double).reshape(3, 8)
        from_c = _multiarray_tests.test_as_c_array(array, 2, 4)
        assert_equal(array[2, 4], from_c)

    def test_3darray(self):
        array = np.arange(24, dtype=np.double).reshape(2, 3, 4)
        from_c = _multiarray_tests.test_as_c_array(array, 1, 2, 3)
        assert_equal(array[1, 2, 3], from_c)


class TestConversion:
    def test_array_scalar_relational_operation(self):
        # All integer
        for dt1 in np.typecodes['AllInteger']:
            assert_(1 > np.array(0, dtype=dt1), f"type {dt1} failed")
            assert_(not 1 < np.array(0, dtype=dt1), f"type {dt1} failed")

            for dt2 in np.typecodes['AllInteger']:
                assert_(np.array(1, dtype=dt1) > np.array(0, dtype=dt2),
                        f"type {dt1} and {dt2} failed")
                assert_(not np.array(1, dtype=dt1) < np.array(0, dtype=dt2),
                        f"type {dt1} and {dt2} failed")

        # Unsigned integers
        for dt1 in 'BHILQP':
            assert_(-1 < np.array(1, dtype=dt1), f"type {dt1} failed")
            assert_(not -1 > np.array(1, dtype=dt1), f"type {dt1} failed")
            assert_(-1 != np.array(1, dtype=dt1), f"type {dt1} failed")

            # Unsigned vs signed
            for dt2 in 'bhilqp':
                assert_(np.array(1, dtype=dt1) > np.array(-1, dtype=dt2),
                        f"type {dt1} and {dt2} failed")
                assert_(not np.array(1, dtype=dt1) < np.array(-1, dtype=dt2),
                        f"type {dt1} and {dt2} failed")
                assert_(np.array(1, dtype=dt1) != np.array(-1, dtype=dt2),
                        f"type {dt1} and {dt2} failed")

        # Signed integers and floats
        for dt1 in 'bhlqp' + np.typecodes['Float']:
            assert_(1 > np.array(-1, dtype=dt1), f"type {dt1} failed")
            assert_(not 1 < np.array(-1, dtype=dt1), f"type {dt1} failed")
            assert_(-1 == np.array(-1, dtype=dt1), f"type {dt1} failed")

            for dt2 in 'bhlqp' + np.typecodes['Float']:
                assert_(np.array(1, dtype=dt1) > np.array(-1, dtype=dt2),
                        f"type {dt1} and {dt2} failed")
                assert_(not np.array(1, dtype=dt1) < np.array(-1, dtype=dt2),
                        f"type {dt1} and {dt2} failed")
                assert_(np.array(-1, dtype=dt1) == np.array(-1, dtype=dt2),
                        f"type {dt1} and {dt2} failed")

    def test_to_bool_scalar(self):
        assert_equal(bool(np.array([False])), False)
        assert_equal(bool(np.array([True])), True)
        assert_equal(bool(np.array([[42]])), True)

    def test_to_bool_scalar_not_convertible(self):

        class NotConvertible:
            def __bool__(self):
                raise NotImplementedError

        assert_raises(NotImplementedError, bool, np.array(NotConvertible()))
        assert_raises(NotImplementedError, bool, np.array([NotConvertible()]))
        if IS_PYSTON:
            pytest.skip("Pyston disables recursion checking")
        if IS_WASM:
            pytest.skip("Pyodide/WASM has limited stack size")

        self_containing = np.array([None])
        self_containing[0] = self_containing

        Error = RecursionError

        assert_raises(Error, bool, self_containing)  # previously stack overflow
        self_containing[0] = None  # resolve circular reference

    def test_to_bool_scalar_size_errors(self):
        with pytest.raises(ValueError, match=".*one element is ambiguous"):
            bool(np.array([1, 2]))

        with pytest.raises(ValueError, match=".*empty array is ambiguous"):
            bool(np.empty((3, 0)))

        with pytest.raises(ValueError, match=".*empty array is ambiguous"):
            bool(np.empty((0,)))

    def test_to_int_scalar(self):
        # gh-9972 means that these aren't always the same
        int_funcs = (int, lambda x: x.__int__())
        for int_func in int_funcs:
            assert_equal(int_func(np.array(0)), 0)
            with assert_warns(DeprecationWarning):
                assert_equal(int_func(np.array([1])), 1)
            with assert_warns(DeprecationWarning):
                assert_equal(int_func(np.array([[42]])), 42)
            assert_raises(TypeError, int_func, np.array([1, 2]))

            # gh-9972
            assert_equal(4, int_func(np.array('4')))
            assert_equal(5, int_func(np.bytes_(b'5')))
            assert_equal(6, int_func(np.str_('6')))

            class NotConvertible:
                def __int__(self):
                    raise NotImplementedError
            assert_raises(NotImplementedError,
                int_func, np.array(NotConvertible()))
            with assert_warns(DeprecationWarning):
                assert_raises(NotImplementedError,
                    int_func, np.array([NotConvertible()]))


class TestWhere:
    def test_basic(self):
        dts = [bool, np.int16, np.int32, np.int64, np.double, np.complex128,
               np.longdouble, np.clongdouble]
        for dt in dts:
            c = np.ones(53, dtype=bool)
            assert_equal(np.where( c, dt(0), dt(1)), dt(0))
            assert_equal(np.where(~c, dt(0), dt(1)), dt(1))
            assert_equal(np.where(True, dt(0), dt(1)), dt(0))
            assert_equal(np.where(False, dt(0), dt(1)), dt(1))
            d = np.ones_like(c).astype(dt)
            e = np.zeros_like(d)
            r = d.astype(dt)
            c[7] = False
            r[7] = e[7]
            assert_equal(np.where(c, e, e), e)
            assert_equal(np.where(c, d, e), r)
            assert_equal(np.where(c, d, e[0]), r)
            assert_equal(np.where(c, d[0], e), r)
            assert_equal(np.where(c[::2], d[::2], e[::2]), r[::2])
            assert_equal(np.where(c[1::2], d[1::2], e[1::2]), r[1::2])
            assert_equal(np.where(c[::3], d[::3], e[::3]), r[::3])
            assert_equal(np.where(c[1::3], d[1::3], e[1::3]), r[1::3])
            assert_equal(np.where(c[::-2], d[::-2], e[::-2]), r[::-2])
            assert_equal(np.where(c[::-3], d[::-3], e[::-3]), r[::-3])
            assert_equal(np.where(c[1::-3], d[1::-3], e[1::-3]), r[1::-3])

    @pytest.mark.skipif(IS_WASM, reason="no wasm fp exception support")
    def test_exotic(self):
        # object
        assert_array_equal(np.where(True, None, None), np.array(None))
        # zero sized
        m = np.array([], dtype=bool).reshape(0, 3)
        b = np.array([], dtype=np.float64).reshape(0, 3)
        assert_array_equal(np.where(m, 0, b), np.array([]).reshape(0, 3))

        # object cast
        d = np.array([-1.34, -0.16, -0.54, -0.31, -0.08, -0.95, 0.000, 0.313,
                      0.547, -0.18, 0.876, 0.236, 1.969, 0.310, 0.699, 1.013,
                      1.267, 0.229, -1.39, 0.487])
        nan = float('NaN')
        e = np.array(['5z', '0l', nan, 'Wz', nan, nan, 'Xq', 'cs', nan, nan,
                     'QN', nan, nan, 'Fd', nan, nan, 'kp', nan, '36', 'i1'],
                     dtype=object)
        m = np.array([0, 0, 1, 0, 1, 1, 0, 0, 1, 1,
                      0, 1, 1, 0, 1, 1, 0, 1, 0, 0], dtype=bool)

        r = e[:]
        r[np.where(m)] = d[np.where(m)]
        assert_array_equal(np.where(m, d, e), r)

        r = e[:]
        r[np.where(~m)] = d[np.where(~m)]
        assert_array_equal(np.where(m, e, d), r)

        assert_array_equal(np.where(m, e, e), e)

        # minimal dtype result with NaN scalar (e.g required by pandas)
        d = np.array([1., 2.], dtype=np.float32)
        e = float('NaN')
        assert_equal(np.where(True, d, e).dtype, np.float32)
        e = float('Infinity')
        assert_equal(np.where(True, d, e).dtype, np.float32)
        e = float('-Infinity')
        assert_equal(np.where(True, d, e).dtype, np.float32)
        # With NEP 50 adopted, the float will overflow here:
        e = 1e150
        with pytest.warns(RuntimeWarning, match="overflow"):
            res = np.where(True, d, e)
        assert res.dtype == np.float32

    def test_ndim(self):
        c = [True, False]
        a = np.zeros((2, 25))
        b = np.ones((2, 25))
        r = np.where(np.array(c)[:, np.newaxis], a, b)
        assert_array_equal(r[0], a[0])
        assert_array_equal(r[1], b[0])

        a = a.T
        b = b.T
        r = np.where(c, a, b)
        assert_array_equal(r[:, 0], a[:, 0])
        assert_array_equal(r[:, 1], b[:, 0])

    def test_dtype_mix(self):
        c = np.array([False, True, False, False, False, False, True, False,
                     False, False, True, False])
        a = np.uint32(1)
        b = np.array([5., 0., 3., 2., -1., -4., 0., -10., 10., 1., 0., 3.],
                      dtype=np.float64)
        r = np.array([5., 1., 3., 2., -1., -4., 1., -10., 10., 1., 1., 3.],
                     dtype=np.float64)
        assert_equal(np.where(c, a, b), r)

        a = a.astype(np.float32)
        b = b.astype(np.int64)
        assert_equal(np.where(c, a, b), r)

        # non bool mask
        c = c.astype(int)
        c[c != 0] = 34242324
        assert_equal(np.where(c, a, b), r)
        # invert
        tmpmask = c != 0
        c[c == 0] = 41247212
        c[tmpmask] = 0
        assert_equal(np.where(c, b, a), r)

    def test_foreign(self):
        c = np.array([False, True, False, False, False, False, True, False,
                     False, False, True, False])
        r = np.array([5., 1., 3., 2., -1., -4., 1., -10., 10., 1., 1., 3.],
                     dtype=np.float64)
        a = np.ones(1, dtype='>i4')
        b = np.array([5., 0., 3., 2., -1., -4., 0., -10., 10., 1., 0., 3.],
                     dtype=np.float64)
        assert_equal(np.where(c, a, b), r)

        b = b.astype('>f8')
        assert_equal(np.where(c, a, b), r)

        a = a.astype('<i4')
        assert_equal(np.where(c, a, b), r)

        c = c.astype('>i4')
        assert_equal(np.where(c, a, b), r)

    def test_error(self):
        c = [True, True]
        a = np.ones((4, 5))
        b = np.ones((5, 5))
        assert_raises(ValueError, np.where, c, a, a)
        assert_raises(ValueError, np.where, c[0], a, b)

    def test_string(self):
        # gh-4778 check strings are properly filled with nulls
        a = np.array("abc")
        b = np.array("x" * 753)
        assert_equal(np.where(True, a, b), "abc")
        assert_equal(np.where(False, b, a), "abc")

        # check native datatype sized strings
        a = np.array("abcd")
        b = np.array("x" * 8)
        assert_equal(np.where(True, a, b), "abcd")
        assert_equal(np.where(False, b, a), "abcd")

    def test_empty_result(self):
        # pass empty where result through an assignment which reads the data of
        # empty arrays, error detectable with valgrind, see gh-8922
        x = np.zeros((1, 1))
        ibad = np.vstack(np.where(x == 99.))
        assert_array_equal(ibad,
                           np.atleast_2d(np.array([[], []], dtype=np.intp)))

    def test_largedim(self):
        # invalid read regression gh-9304
        shape = [10, 2, 3, 4, 5, 6]
        np.random.seed(2)
        array = np.random.rand(*shape)

        for i in range(10):
            benchmark = array.nonzero()
            result = array.nonzero()
            assert_array_equal(benchmark, result)

    def test_kwargs(self):
        a = np.zeros(1)
        with assert_raises(TypeError):
            np.where(a, x=a, y=a)


if not IS_PYPY:
    # sys.getsizeof() is not valid on PyPy
    class TestSizeOf:

        def test_empty_array(self):
            x = np.array([])
            assert_(sys.getsizeof(x) > 0)

        def check_array(self, dtype):
            elem_size = dtype(0).itemsize

            for length in [10, 50, 100, 500]:
                x = np.arange(length, dtype=dtype)
                assert_(sys.getsizeof(x) > length * elem_size)

        def test_array_int32(self):
            self.check_array(np.int32)

        def test_array_int64(self):
            self.check_array(np.int64)

        def test_array_float32(self):
            self.check_array(np.float32)

        def test_array_float64(self):
            self.check_array(np.float64)

        def test_view(self):
            d = np.ones(100)
            assert_(sys.getsizeof(d[...]) < sys.getsizeof(d))

        def test_reshape(self):
            d = np.ones(100)
            assert_(sys.getsizeof(d) < sys.getsizeof(d.reshape(100, 1, 1).copy()))

        @_no_tracing
        def test_resize(self):
            d = np.ones(100)
            old = sys.getsizeof(d)
            d.resize(50)
            assert_(old > sys.getsizeof(d))
            d.resize(150)
            assert_(old < sys.getsizeof(d))

        @pytest.mark.parametrize("dtype", ["u4,f4", "u4,O"])
        def test_resize_structured(self, dtype):
            a = np.array([(0, 0.0) for i in range(5)], dtype=dtype)
            a.resize(1000)
            assert_array_equal(a, np.zeros(1000, dtype=dtype))

        def test_error(self):
            d = np.ones(100)
            assert_raises(TypeError, d.__sizeof__, "a")


class TestHashing:

    def test_arrays_not_hashable(self):
        x = np.ones(3)
        assert_raises(TypeError, hash, x)

    def test_collections_hashable(self):
        x = np.array([])
        assert_(not isinstance(x, collections.abc.Hashable))


class TestArrayPriority:
    # This will go away when __array_priority__ is settled, meanwhile
    # it serves to check unintended changes.
    op = operator
    binary_ops = [
        op.pow, op.add, op.sub, op.mul, op.floordiv, op.truediv, op.mod,
        op.and_, op.or_, op.xor, op.lshift, op.rshift, op.mod, op.gt,
        op.ge, op.lt, op.le, op.ne, op.eq
        ]

    class Foo(np.ndarray):
        __array_priority__ = 100.

        def __new__(cls, *args, **kwargs):
            return np.array(*args, **kwargs).view(cls)

    class Bar(np.ndarray):
        __array_priority__ = 101.

        def __new__(cls, *args, **kwargs):
            return np.array(*args, **kwargs).view(cls)

    class Other:
        __array_priority__ = 1000.

        def _all(self, other):
            return self.__class__()

        __add__ = __radd__ = _all
        __sub__ = __rsub__ = _all
        __mul__ = __rmul__ = _all
        __pow__ = __rpow__ = _all
        __mod__ = __rmod__ = _all
        __truediv__ = __rtruediv__ = _all
        __floordiv__ = __rfloordiv__ = _all
        __and__ = __rand__ = _all
        __xor__ = __rxor__ = _all
        __or__ = __ror__ = _all
        __lshift__ = __rlshift__ = _all
        __rshift__ = __rrshift__ = _all
        __eq__ = _all
        __ne__ = _all
        __gt__ = _all
        __ge__ = _all
        __lt__ = _all
        __le__ = _all

    def test_ndarray_subclass(self):
        a = np.array([1, 2])
        b = self.Bar([1, 2])
        for f in self.binary_ops:
            msg = repr(f)
            assert_(isinstance(f(a, b), self.Bar), msg)
            assert_(isinstance(f(b, a), self.Bar), msg)

    def test_ndarray_other(self):
        a = np.array([1, 2])
        b = self.Other()
        for f in self.binary_ops:
            msg = repr(f)
            assert_(isinstance(f(a, b), self.Other), msg)
            assert_(isinstance(f(b, a), self.Other), msg)

    def test_subclass_subclass(self):
        a = self.Foo([1, 2])
        b = self.Bar([1, 2])
        for f in self.binary_ops:
            msg = repr(f)
            assert_(isinstance(f(a, b), self.Bar), msg)
            assert_(isinstance(f(b, a), self.Bar), msg)

    def test_subclass_other(self):
        a = self.Foo([1, 2])
        b = self.Other()
        for f in self.binary_ops:
            msg = repr(f)
            assert_(isinstance(f(a, b), self.Other), msg)
            assert_(isinstance(f(b, a), self.Other), msg)


class TestBytestringArrayNonzero:

    def test_empty_bstring_array_is_falsey(self):
        assert_(not np.array([''], dtype=str))

    def test_whitespace_bstring_array_is_truthy(self):
        a = np.array(['spam'], dtype=str)
        a[0] = '  \0\0'
        assert_(a)

    def test_all_null_bstring_array_is_falsey(self):
        a = np.array(['spam'], dtype=str)
        a[0] = '\0\0\0\0'
        assert_(not a)

    def test_null_inside_bstring_array_is_truthy(self):
        a = np.array(['spam'], dtype=str)
        a[0] = ' \0 \0'
        assert_(a)


class TestUnicodeEncoding:
    """
    Tests for encoding related bugs, such as UCS2 vs UCS4, round-tripping
    issues, etc
    """
    def test_round_trip(self):
        """ Tests that GETITEM, SETITEM, and PyArray_Scalar roundtrip """
        # gh-15363
        arr = np.zeros(shape=(), dtype="U1")
        for i in range(1, sys.maxunicode + 1):
            expected = chr(i)
            arr[()] = expected
            assert arr[()] == expected
            assert arr.item() == expected

    def test_assign_scalar(self):
        # gh-3258
        l = np.array(['aa', 'bb'])
        l[:] = np.str_('cc')
        assert_equal(l, ['cc', 'cc'])

    def test_fill_scalar(self):
        # gh-7227
        l = np.array(['aa', 'bb'])
        l.fill(np.str_('cc'))
        assert_equal(l, ['cc', 'cc'])


class TestUnicodeArrayNonzero:

    def test_empty_ustring_array_is_falsey(self):
        assert_(not np.array([''], dtype=np.str_))

    def test_whitespace_ustring_array_is_truthy(self):
        a = np.array(['eggs'], dtype=np.str_)
        a[0] = '  \0\0'
        assert_(a)

    def test_all_null_ustring_array_is_falsey(self):
        a = np.array(['eggs'], dtype=np.str_)
        a[0] = '\0\0\0\0'
        assert_(not a)

    def test_null_inside_ustring_array_is_truthy(self):
        a = np.array(['eggs'], dtype=np.str_)
        a[0] = ' \0 \0'
        assert_(a)


class TestFormat:

    def test_0d(self):
        a = np.array(np.pi)
        assert_equal(f'{a:0.3g}', '3.14')
        assert_equal(f'{a[()]:0.3g}', '3.14')

    def test_1d_no_format(self):
        a = np.array([np.pi])
        assert_equal(f'{a}', str(a))

    def test_1d_format(self):
        # until gh-5543, ensure that the behaviour matches what it used to be
        a = np.array([np.pi])
        assert_raises(TypeError, '{:30}'.format, a)


from numpy.testing import IS_PYPY


class TestCTypes:

    def test_ctypes_is_available(self):
        test_arr = np.array([[1, 2, 3], [4, 5, 6]])

        assert_equal(ctypes, test_arr.ctypes._ctypes)
        assert_equal(tuple(test_arr.ctypes.shape), (2, 3))

    def test_ctypes_is_not_available(self):
        from numpy._core import _internal
        _internal.ctypes = None
        try:
            test_arr = np.array([[1, 2, 3], [4, 5, 6]])

            assert_(isinstance(test_arr.ctypes._ctypes,
                               _internal._missing_ctypes))
            assert_equal(tuple(test_arr.ctypes.shape), (2, 3))
        finally:
            _internal.ctypes = ctypes

    def _make_readonly(x):
        x.flags.writeable = False
        return x

    @pytest.mark.parametrize('arr', [
        np.array([1, 2, 3]),
        np.array([['one', 'two'], ['three', 'four']]),
        np.array((1, 2), dtype='i4,i4'),
        np.zeros((2,), dtype=np.dtype({
                "formats": ['<i4', '<i4'],
                "names": ['a', 'b'],
                "offsets": [0, 2],
                "itemsize": 6
            })
        ),
        np.array([None], dtype=object),
        np.array([]),
        np.empty((0, 0)),
        _make_readonly(np.array([1, 2, 3])),
    ], ids=[
        '1d',
        '2d',
        'structured',
        'overlapping',
        'object',
        'empty',
        'empty-2d',
        'readonly'
    ])
    def test_ctypes_data_as_holds_reference(self, arr):
        # gh-9647
        # create a copy to ensure that pytest does not mess with the refcounts
        arr = arr.copy()

        arr_ref = weakref.ref(arr)

        ctypes_ptr = arr.ctypes.data_as(ctypes.c_void_p)

        # `ctypes_ptr` should hold onto `arr`
        del arr
        break_cycles()
        assert_(arr_ref() is not None, "ctypes pointer did not hold onto a reference")

        # but when the `ctypes_ptr` object dies, so should `arr`
        del ctypes_ptr
        if IS_PYPY:
            # Pypy does not recycle arr objects immediately. Trigger gc to
            # release arr. Cpython uses refcounts. An explicit call to gc
            # should not be needed here.
            break_cycles()
        assert_(arr_ref() is None, "unknowable whether ctypes pointer holds a reference")

    def test_ctypes_as_parameter_holds_reference(self):
        arr = np.array([None]).copy()

        arr_ref = weakref.ref(arr)

        ctypes_ptr = arr.ctypes._as_parameter_

        # `ctypes_ptr` should hold onto `arr`
        del arr
        break_cycles()
        assert_(arr_ref() is not None, "ctypes pointer did not hold onto a reference")

        # but when the `ctypes_ptr` object dies, so should `arr`
        del ctypes_ptr
        if IS_PYPY:
            break_cycles()
        assert_(arr_ref() is None, "unknowable whether ctypes pointer holds a reference")


class TestWritebackIfCopy:
    # all these tests use the WRITEBACKIFCOPY mechanism
    def test_argmax_with_out(self):
        mat = np.eye(5)
        out = np.empty(5, dtype='i2')
        res = np.argmax(mat, 0, out=out)
        assert_equal(res, range(5))

    def test_argmin_with_out(self):
        mat = -np.eye(5)
        out = np.empty(5, dtype='i2')
        res = np.argmin(mat, 0, out=out)
        assert_equal(res, range(5))

    def test_insert_noncontiguous(self):
        a = np.arange(6).reshape(2, 3).T  # force non-c-contiguous
        # uses arr_insert
        np.place(a, a > 2, [44, 55])
        assert_equal(a, np.array([[0, 44], [1, 55], [2, 44]]))
        # hit one of the failing paths
        assert_raises(ValueError, np.place, a, a > 20, [])

    def test_put_noncontiguous(self):
        a = np.arange(6).reshape(2, 3).T  # force non-c-contiguous
        np.put(a, [0, 2], [44, 55])
        assert_equal(a, np.array([[44, 3], [55, 4], [2, 5]]))

    def test_putmask_noncontiguous(self):
        a = np.arange(6).reshape(2, 3).T  # force non-c-contiguous
        # uses arr_putmask
        np.putmask(a, a > 2, a**2)
        assert_equal(a, np.array([[0, 9], [1, 16], [2, 25]]))

    def test_take_mode_raise(self):
        a = np.arange(6, dtype='int')
        out = np.empty(2, dtype='int')
        np.take(a, [0, 2], out=out, mode='raise')
        assert_equal(out, np.array([0, 2]))

    def test_choose_mod_raise(self):
        a = np.array([[1, 0, 1], [0, 1, 0], [1, 0, 1]])
        out = np.empty((3, 3), dtype='int')
        choices = [-10, 10]
        np.choose(a, choices, out=out, mode='raise')
        assert_equal(out, np.array([[ 10, -10,  10],
                                    [-10,  10, -10],
                                    [ 10, -10,  10]]))

    def test_flatiter__array__(self):
        a = np.arange(9).reshape(3, 3)
        b = a.T.flat
        c = b.__array__()
        # triggers the WRITEBACKIFCOPY resolution, assuming refcount semantics
        del c

    def test_dot_out(self):
        # if HAVE_CBLAS, will use WRITEBACKIFCOPY
        a = np.arange(9, dtype=float).reshape(3, 3)
        b = np.dot(a, a, out=a)
        assert_equal(b, np.array([[15, 18, 21], [42, 54, 66], [69, 90, 111]]))

    def test_view_assign(self):
        from numpy._core._multiarray_tests import (
            npy_create_writebackifcopy,
            npy_resolve,
        )

        arr = np.arange(9).reshape(3, 3).T
        arr_wb = npy_create_writebackifcopy(arr)
        assert_(arr_wb.flags.writebackifcopy)
        assert_(arr_wb.base is arr)
        arr_wb[...] = -100
        npy_resolve(arr_wb)
        # arr changes after resolve, even though we assigned to arr_wb
        assert_equal(arr, -100)
        # after resolve, the two arrays no longer reference each other
        assert_(arr_wb.ctypes.data != 0)
        assert_equal(arr_wb.base, None)
        # assigning to arr_wb does not get transferred to arr
        arr_wb[...] = 100
        assert_equal(arr, -100)

    @pytest.mark.leaks_references(
            reason="increments self in dealloc; ignore since deprecated path.")
    def test_dealloc_warning(self):
        with suppress_warnings() as sup:
            sup.record(RuntimeWarning)
            arr = np.arange(9).reshape(3, 3)
            v = arr.T
            _multiarray_tests.npy_abuse_writebackifcopy(v)
            assert len(sup.log) == 1

    def test_view_discard_refcount(self):
        from numpy._core._multiarray_tests import (
            npy_create_writebackifcopy,
            npy_discard,
        )

        arr = np.arange(9).reshape(3, 3).T
        orig = arr.copy()
        if HAS_REFCOUNT:
            arr_cnt = sys.getrefcount(arr)
        arr_wb = npy_create_writebackifcopy(arr)
        assert_(arr_wb.flags.writebackifcopy)
        assert_(arr_wb.base is arr)
        arr_wb[...] = -100
        npy_discard(arr_wb)
        # arr remains unchanged after discard
        assert_equal(arr, orig)
        # after discard, the two arrays no longer reference each other
        assert_(arr_wb.ctypes.data != 0)
        assert_equal(arr_wb.base, None)
        if HAS_REFCOUNT:
            assert_equal(arr_cnt, sys.getrefcount(arr))
        # assigning to arr_wb does not get transferred to arr
        arr_wb[...] = 100
        assert_equal(arr, orig)


class TestArange:
    def test_infinite(self):
        assert_raises_regex(
            ValueError, "size exceeded",
            np.arange, 0, np.inf
        )

    def test_nan_step(self):
        assert_raises_regex(
            ValueError, "cannot compute length",
            np.arange, 0, 1, np.nan
        )

    def test_zero_step(self):
        assert_raises(ZeroDivisionError, np.arange, 0, 10, 0)
        assert_raises(ZeroDivisionError, np.arange, 0.0, 10.0, 0.0)

        # empty range
        assert_raises(ZeroDivisionError, np.arange, 0, 0, 0)
        assert_raises(ZeroDivisionError, np.arange, 0.0, 0.0, 0.0)

    def test_require_range(self):
        assert_raises(TypeError, np.arange)
        assert_raises(TypeError, np.arange, step=3)
        assert_raises(TypeError, np.arange, dtype='int64')
        assert_raises(TypeError, np.arange, start=4)

    def test_start_stop_kwarg(self):
        keyword_stop = np.arange(stop=3)
        keyword_zerotostop = np.arange(0, stop=3)
        keyword_start_stop = np.arange(start=3, stop=9)

        assert len(keyword_stop) == 3
        assert len(keyword_zerotostop) == 3
        assert len(keyword_start_stop) == 6
        assert_array_equal(keyword_stop, keyword_zerotostop)

    def test_arange_booleans(self):
        # Arange makes some sense for booleans and works up to length 2.
        # But it is weird since `arange(2, 4, dtype=bool)` works.
        # Arguably, much or all of this could be deprecated/removed.
        res = np.arange(False, dtype=bool)
        assert_array_equal(res, np.array([], dtype="bool"))

        res = np.arange(True, dtype="bool")
        assert_array_equal(res, [False])

        res = np.arange(2, dtype="bool")
        assert_array_equal(res, [False, True])

        # This case is especially weird, but drops out without special case:
        res = np.arange(6, 8, dtype="bool")
        assert_array_equal(res, [True, True])

        with pytest.raises(TypeError):
            np.arange(3, dtype="bool")

    @pytest.mark.parametrize("dtype", ["S3", "U", "5i"])
    def test_rejects_bad_dtypes(self, dtype):
        dtype = np.dtype(dtype)
        DType_name = re.escape(str(type(dtype)))
        with pytest.raises(TypeError,
                match=rf"arange\(\) not supported for inputs .* {DType_name}"):
            np.arange(2, dtype=dtype)

    def test_rejects_strings(self):
        # Explicitly test error for strings which may call "b" - "a":
        DType_name = re.escape(str(type(np.array("a").dtype)))
        with pytest.raises(TypeError,
                match=rf"arange\(\) not supported for inputs .* {DType_name}"):
            np.arange("a", "b")

    def test_byteswapped(self):
        res_be = np.arange(1, 1000, dtype=">i4")
        res_le = np.arange(1, 1000, dtype="<i4")
        assert res_be.dtype == ">i4"
        assert res_le.dtype == "<i4"
        assert_array_equal(res_le, res_be)

    @pytest.mark.parametrize("which", [0, 1, 2])
    def test_error_paths_and_promotion(self, which):
        args = [0, 1, 2]  # start, stop, and step
        args[which] = np.float64(2.)  # should ensure float64 output

        assert np.arange(*args).dtype == np.float64

        # Cover stranger error path, test only to achieve code coverage!
        args[which] = [None, []]
        with pytest.raises(ValueError):
            # Fails discovering start dtype
            np.arange(*args)

    def test_dtype_attribute_ignored(self):
        # Until 2.3 this would raise a DeprecationWarning
        class dt:
            dtype = "f8"

        class vdt(np.void):
            dtype = "f,f"

        assert_raises(ValueError, np.dtype, dt)
        assert_raises(ValueError, np.dtype, dt())
        assert_raises(ValueError, np.dtype, vdt)
        assert_raises(ValueError, np.dtype, vdt(1))


class TestDTypeCoercionForbidden:
    forbidden_types = [
        # The builtin scalar super types:
        np.generic, np.flexible, np.number,
        np.inexact, np.floating, np.complexfloating,
        np.integer, np.unsignedinteger, np.signedinteger,
        # character is a deprecated S1 special case:
        np.character,
    ]

    def test_dtype_coercion(self):
        for scalar_type in self.forbidden_types:
            assert_raises(TypeError, np.dtype, args=(scalar_type,))

    def test_array_construction(self):
        for scalar_type in self.forbidden_types:
            assert_raises(TypeError, np.array, args=([], scalar_type,))

    def test_not_deprecated(self):
        # All specific types work
        for group in np._core.sctypes.values():
            for scalar_type in group:
                np.dtype(scalar_type)

        for scalar_type in [type, dict, list, tuple]:
            # Typical python types are coerced to object currently:
            np.dtype(scalar_type)


class TestDateTimeCreationTuple:
    @pytest.mark.parametrize("cls", [np.datetime64, np.timedelta64])
    def test_dt_tuple(self, cls):
        # two valid uses - (unit, num) and (unit, num, den, None)
        cls(1, ('ms', 2))
        cls(1, ('ms', 2, 1, None))

        # trying to use the event argument, removed in 1.7.0
        # it used to be a uint8
        assert_raises(TypeError, cls, args=(1, ('ms', 2, 'event')))
        assert_raises(TypeError, cls, args=(1, ('ms', 2, 63)))
        assert_raises(TypeError, cls, args=(1, ('ms', 2, 1, 'event')))
        assert_raises(TypeError, cls, args=(1, ('ms', 2, 1, 63)))


class TestArrayFinalize:
    """ Tests __array_finalize__ """

    def test_receives_base(self):
        # gh-11237
        class SavesBase(np.ndarray):
            def __array_finalize__(self, obj):
                self.saved_base = self.base

        a = np.array(1).view(SavesBase)
        assert_(a.saved_base is a.base)

    def test_bad_finalize1(self):
        class BadAttributeArray(np.ndarray):
            @property
            def __array_finalize__(self):
                raise RuntimeError("boohoo!")

        with pytest.raises(TypeError, match="not callable"):
            np.arange(10).view(BadAttributeArray)

    def test_bad_finalize2(self):
        class BadAttributeArray(np.ndarray):
            def __array_finalize__(self):
                raise RuntimeError("boohoo!")

        with pytest.raises(TypeError, match="takes 1 positional"):
            np.arange(10).view(BadAttributeArray)

    def test_bad_finalize3(self):
        class BadAttributeArray(np.ndarray):
            def __array_finalize__(self, obj):
                raise RuntimeError("boohoo!")

        with pytest.raises(RuntimeError, match="boohoo!"):
            np.arange(10).view(BadAttributeArray)

    def test_lifetime_on_error(self):
        # gh-11237
        class RaisesInFinalize(np.ndarray):
            def __array_finalize__(self, obj):
                # crash, but keep this object alive
                raise Exception(self)

        # a plain object can't be weakref'd
        class Dummy:
            pass

        # get a weak reference to an object within an array
        obj_arr = np.array(Dummy())
        obj_ref = weakref.ref(obj_arr[()])

        # get an array that crashed in __array_finalize__
        with assert_raises(Exception) as e:
            obj_arr.view(RaisesInFinalize)

        obj_subarray = e.exception.args[0]
        del e
        assert_(isinstance(obj_subarray, RaisesInFinalize))

        # reference should still be held by obj_arr
        break_cycles()
        assert_(obj_ref() is not None, "object should not already be dead")

        del obj_arr
        break_cycles()
        assert_(obj_ref() is not None, "obj_arr should not hold the last reference")

        del obj_subarray
        break_cycles()
        assert_(obj_ref() is None, "no references should remain")

    def test_can_use_super(self):
        class SuperFinalize(np.ndarray):
            def __array_finalize__(self, obj):
                self.saved_result = super().__array_finalize__(obj)

        a = np.array(1).view(SuperFinalize)
        assert_(a.saved_result is None)


def test_orderconverter_with_nonASCII_unicode_ordering():
    # gh-7475
    a = np.arange(5)
    assert_raises(ValueError, a.flatten, order='\xe2')


def test_equal_override():
    # gh-9153: ndarray.__eq__ uses special logic for structured arrays, which
    # did not respect overrides with __array_priority__ or __array_ufunc__.
    # The PR fixed this for __array_priority__ and __array_ufunc__ = None.
    class MyAlwaysEqual:
        def __eq__(self, other):
            return "eq"

        def __ne__(self, other):
            return "ne"

    class MyAlwaysEqualOld(MyAlwaysEqual):
        __array_priority__ = 10000

    class MyAlwaysEqualNew(MyAlwaysEqual):
        __array_ufunc__ = None

    array = np.array([(0, 1), (2, 3)], dtype='i4,i4')
    for my_always_equal_cls in MyAlwaysEqualOld, MyAlwaysEqualNew:
        my_always_equal = my_always_equal_cls()
        assert_equal(my_always_equal == array, 'eq')
        assert_equal(array == my_always_equal, 'eq')
        assert_equal(my_always_equal != array, 'ne')
        assert_equal(array != my_always_equal, 'ne')


@pytest.mark.parametrize("op", [operator.eq, operator.ne])
@pytest.mark.parametrize(["dt1", "dt2"], [
        ([("f", "i")], [("f", "i")]),  # structured comparison (successful)
        ("M8", "d"),  # impossible comparison: result is all True or False
        ("d", "d"),  # valid comparison
        ])
def test_equal_subclass_no_override(op, dt1, dt2):
    # Test how the three different possible code-paths deal with subclasses

    class MyArr(np.ndarray):
        called_wrap = 0

        def __array_wrap__(self, new, context=None, return_scalar=False):
            type(self).called_wrap += 1
            return super().__array_wrap__(new, context, return_scalar)

    numpy_arr = np.zeros(5, dtype=dt1)
    my_arr = np.zeros(5, dtype=dt2).view(MyArr)

    assert type(op(numpy_arr, my_arr)) is MyArr
    assert type(op(my_arr, numpy_arr)) is MyArr
    # We expect 2 calls (more if there were more fields):
    assert MyArr.called_wrap == 2


@pytest.mark.parametrize(["dt1", "dt2"], [
        ("M8[ns]", "d"),
        ("M8[s]", "l"),
        ("m8[ns]", "d"),
        # Missing: ("m8[ns]", "l") as timedelta currently promotes ints
        ("M8[s]", "m8[s]"),
        ("S5", "U5"),
        # Structured/void dtypes have explicit paths not tested here.
])
def test_no_loop_gives_all_true_or_false(dt1, dt2):
    # Make sure they broadcast to test result shape, use random values, since
    # the actual value should be ignored
    arr1 = np.random.randint(5, size=100).astype(dt1)
    arr2 = np.random.randint(5, size=99)[:, np.newaxis].astype(dt2)

    res = arr1 == arr2
    assert res.shape == (99, 100)
    assert res.dtype == bool
    assert not res.any()

    res = arr1 != arr2
    assert res.shape == (99, 100)
    assert res.dtype == bool
    assert res.all()

    # incompatible shapes raise though
    arr2 = np.random.randint(5, size=99).astype(dt2)
    with pytest.raises(ValueError):
        arr1 == arr2

    with pytest.raises(ValueError):
        arr1 != arr2

    # Basic test with another operation:
    with pytest.raises(np._core._exceptions._UFuncNoLoopError):
        arr1 > arr2


@pytest.mark.parametrize("op", [
        operator.eq, operator.ne, operator.le, operator.lt, operator.ge,
        operator.gt])
def test_comparisons_forwards_error(op):
    class NotArray:
        def __array__(self, dtype=None, copy=None):
            raise TypeError("run you fools")

    with pytest.raises(TypeError, match="run you fools"):
        op(np.arange(2), NotArray())

    with pytest.raises(TypeError, match="run you fools"):
        op(NotArray(), np.arange(2))


def test_richcompare_scalar_boolean_singleton_return():
    # These are currently guaranteed to be the boolean numpy singletons
    assert (np.array(0) == "a") is np.bool_(False)
    assert (np.array(0) != "a") is np.bool_(True)
    assert (np.int16(0) == "a") is np.bool_(False)
    assert (np.int16(0) != "a") is np.bool_(True)


@pytest.mark.parametrize("op", [
        operator.eq, operator.ne, operator.le, operator.lt, operator.ge,
        operator.gt])
def test_ragged_comparison_fails(op):
    # This needs to convert the internal array to True/False, which fails:
    a = np.array([1, np.array([1, 2, 3])], dtype=object)
    b = np.array([1, np.array([1, 2, 3])], dtype=object)

    with pytest.raises(ValueError, match="The truth value.*ambiguous"):
        op(a, b)


@pytest.mark.parametrize(
    ["fun", "npfun"],
    [
        (_multiarray_tests.npy_cabs, np.absolute),
        (_multiarray_tests.npy_carg, np.angle)
    ]
)
@pytest.mark.parametrize("x", [1, np.inf, -np.inf, np.nan])
@pytest.mark.parametrize("y", [1, np.inf, -np.inf, np.nan])
@pytest.mark.parametrize("test_dtype", np.complexfloating.__subclasses__())
def test_npymath_complex(fun, npfun, x, y, test_dtype):
    # Smoketest npymath functions
    z = test_dtype(complex(x, y))
    with np.errstate(invalid='ignore'):
        # Fallback implementations may emit a warning for +-inf (see gh-24876):
        #     RuntimeWarning: invalid value encountered in absolute
        got = fun(z)
        expected = npfun(z)
        assert_allclose(got, expected)


def test_npymath_real():
    # Smoketest npymath functions
    from numpy._core._multiarray_tests import (
        npy_cosh,
        npy_log10,
        npy_sinh,
        npy_tan,
        npy_tanh,
    )

    funcs = {npy_log10: np.log10,
             npy_cosh: np.cosh,
             npy_sinh: np.sinh,
             npy_tan: np.tan,
             npy_tanh: np.tanh}
    vals = (1, np.inf, -np.inf, np.nan)
    types = (np.float32, np.float64, np.longdouble)

    with np.errstate(all='ignore'):
        for fun, npfun in funcs.items():
            for x, t in itertools.product(vals, types):
                z = t(x)
                got = fun(z)
                expected = npfun(z)
                assert_allclose(got, expected)

def test_uintalignment_and_alignment():
    # alignment code needs to satisfy these requirements:
    #  1. numpy structs match C struct layout
    #  2. ufuncs/casting is safe wrt to aligned access
    #  3. copy code is safe wrt to "uint alidned" access
    #
    # Complex types are the main problem, whose alignment may not be the same
    # as their "uint alignment".
    #
    # This test might only fail on certain platforms, where uint64 alignment is
    # not equal to complex64 alignment. The second 2 tests will only fail
    # for DEBUG=1.

    d1 = np.dtype('u1,c8', align=True)
    d2 = np.dtype('u4,c8', align=True)
    d3 = np.dtype({'names': ['a', 'b'], 'formats': ['u1', d1]}, align=True)

    assert_equal(np.zeros(1, dtype=d1)['f1'].flags['ALIGNED'], True)
    assert_equal(np.zeros(1, dtype=d2)['f1'].flags['ALIGNED'], True)
    assert_equal(np.zeros(1, dtype='u1,c8')['f1'].flags['ALIGNED'], False)

    # check that C struct matches numpy struct size
    s = _multiarray_tests.get_struct_alignments()
    for d, (alignment, size) in zip([d1, d2, d3], s):
        assert_equal(d.alignment, alignment)
        assert_equal(d.itemsize, size)

    # check that ufuncs don't complain in debug mode
    # (this is probably OK if the aligned flag is true above)
    src = np.zeros((2, 2), dtype=d1)['f1']  # 4-byte aligned, often
    np.exp(src)  # assert fails?

    # check that copy code doesn't complain in debug mode
    dst = np.zeros((2, 2), dtype='c8')
    dst[:, 1] = src[:, 1]  # assert in lowlevel_strided_loops fails?

class TestAlignment:
    # adapted from scipy._lib.tests.test__util.test__aligned_zeros
    # Checks that unusual memory alignments don't trip up numpy.

    def check(self, shape, dtype, order, align):
        err_msg = repr((shape, dtype, order, align))
        x = _aligned_zeros(shape, dtype, order, align=align)
        if align is None:
            align = np.dtype(dtype).alignment
        assert_equal(x.__array_interface__['data'][0] % align, 0)
        if hasattr(shape, '__len__'):
            assert_equal(x.shape, shape, err_msg)
        else:
            assert_equal(x.shape, (shape,), err_msg)
        assert_equal(x.dtype, dtype)
        if order == "C":
            assert_(x.flags.c_contiguous, err_msg)
        elif order == "F":
            if x.size > 0:
                assert_(x.flags.f_contiguous, err_msg)
        elif order is None:
            assert_(x.flags.c_contiguous, err_msg)
        else:
            raise ValueError

    def test_various_alignments(self):
        for align in [1, 2, 3, 4, 8, 12, 16, 32, 64, None]:
            for n in [0, 1, 3, 11]:
                for order in ["C", "F", None]:
                    for dtype in list(np.typecodes["All"]) + ['i4,i4,i4']:
                        if dtype == 'O':
                            # object dtype can't be misaligned
                            continue
                        for shape in [n, (1, 2, 3, n)]:
                            self.check(shape, np.dtype(dtype), order, align)

    def test_strided_loop_alignments(self):
        # particularly test that complex64 and float128 use right alignment
        # code-paths, since these are particularly problematic. It is useful to
        # turn on USE_DEBUG for this test, so lowlevel-loop asserts are run.
        for align in [1, 2, 4, 8, 12, 16, None]:
            xf64 = _aligned_zeros(3, np.float64)

            xc64 = _aligned_zeros(3, np.complex64, align=align)
            xf128 = _aligned_zeros(3, np.longdouble, align=align)

            # test casting, both to and from misaligned
            with suppress_warnings() as sup:
                sup.filter(ComplexWarning, "Casting complex values")
                xc64.astype('f8')
            xf64.astype(np.complex64)
            test = xc64 + xf64

            xf128.astype('f8')
            xf64.astype(np.longdouble)
            test = xf128 + xf64

            test = xf128 + xc64

            # test copy, both to and from misaligned
            # contig copy
            xf64[:] = xf64.copy()
            xc64[:] = xc64.copy()
            xf128[:] = xf128.copy()
            # strided copy
            xf64[::2] = xf64[::2].copy()
            xc64[::2] = xc64[::2].copy()
            xf128[::2] = xf128[::2].copy()

def test_getfield():
    a = np.arange(32, dtype='uint16')
    if sys.byteorder == 'little':
        i = 0
        j = 1
    else:
        i = 1
        j = 0
    b = a.getfield('int8', i)
    assert_equal(b, a)
    b = a.getfield('int8', j)
    assert_equal(b, 0)
    pytest.raises(ValueError, a.getfield, 'uint8', -1)
    pytest.raises(ValueError, a.getfield, 'uint8', 16)
    pytest.raises(ValueError, a.getfield, 'uint64', 0)


class TestViewDtype:
    """
    Verify that making a view of a non-contiguous array works as expected.
    """
    def test_smaller_dtype_multiple(self):
        # x is non-contiguous
        x = np.arange(10, dtype='<i4')[::2]
        with pytest.raises(ValueError,
                           match='the last axis must be contiguous'):
            x.view('<i2')
        expected = [[0, 0], [2, 0], [4, 0], [6, 0], [8, 0]]
        assert_array_equal(x[:, np.newaxis].view('<i2'), expected)

    def test_smaller_dtype_not_multiple(self):
        # x is non-contiguous
        x = np.arange(5, dtype='<i4')[::2]

        with pytest.raises(ValueError,
                           match='the last axis must be contiguous'):
            x.view('S3')
        with pytest.raises(ValueError,
                           match='When changing to a smaller dtype'):
            x[:, np.newaxis].view('S3')

        # Make sure the problem is because of the dtype size
        expected = [[b''], [b'\x02'], [b'\x04']]
        assert_array_equal(x[:, np.newaxis].view('S4'), expected)

    def test_larger_dtype_multiple(self):
        # x is non-contiguous in the first dimension, contiguous in the last
        x = np.arange(20, dtype='<i2').reshape(10, 2)[::2, :]
        expected = np.array([[65536], [327684], [589832],
                             [851980], [1114128]], dtype='<i4')
        assert_array_equal(x.view('<i4'), expected)

    def test_larger_dtype_not_multiple(self):
        # x is non-contiguous in the first dimension, contiguous in the last
        x = np.arange(20, dtype='<i2').reshape(10, 2)[::2, :]
        with pytest.raises(ValueError,
                           match='When changing to a larger dtype'):
            x.view('S3')
        # Make sure the problem is because of the dtype size
        expected = [[b'\x00\x00\x01'], [b'\x04\x00\x05'], [b'\x08\x00\t'],
                    [b'\x0c\x00\r'], [b'\x10\x00\x11']]
        assert_array_equal(x.view('S4'), expected)

    def test_f_contiguous(self):
        # x is F-contiguous
        x = np.arange(4 * 3, dtype='<i4').reshape(4, 3).T
        with pytest.raises(ValueError,
                           match='the last axis must be contiguous'):
            x.view('<i2')

    def test_non_c_contiguous(self):
        # x is contiguous in axis=-1, but not C-contiguous in other axes
        x = np.arange(2 * 3 * 4, dtype='i1').\
                    reshape(2, 3, 4).transpose(1, 0, 2)
        expected = [[[256, 770], [3340, 3854]],
                    [[1284, 1798], [4368, 4882]],
                    [[2312, 2826], [5396, 5910]]]
        assert_array_equal(x.view('<i2'), expected)


@pytest.mark.xfail(check_support_sve(), reason="gh-22982")
# Test various array sizes that hit different code paths in quicksort-avx512
@pytest.mark.parametrize("N", np.arange(1, 512))
@pytest.mark.parametrize("dtype", ['e', 'f', 'd'])
def test_sort_float(N, dtype):
    # Regular data with nan sprinkled
    np.random.seed(42)
    arr = -0.5 + np.random.sample(N).astype(dtype)
    arr[np.random.choice(arr.shape[0], 3)] = np.nan
    assert_equal(np.sort(arr, kind='quick'), np.sort(arr, kind='heap'))

    # (2) with +INF
    infarr = np.inf * np.ones(N, dtype=dtype)
    infarr[np.random.choice(infarr.shape[0], 5)] = -1.0
    assert_equal(np.sort(infarr, kind='quick'), np.sort(infarr, kind='heap'))

    # (3) with -INF
    neginfarr = -np.inf * np.ones(N, dtype=dtype)
    neginfarr[np.random.choice(neginfarr.shape[0], 5)] = 1.0
    assert_equal(np.sort(neginfarr, kind='quick'),
                 np.sort(neginfarr, kind='heap'))

    # (4) with +/-INF
    infarr = np.inf * np.ones(N, dtype=dtype)
    infarr[np.random.choice(infarr.shape[0], (int)(N / 2))] = -np.inf
    assert_equal(np.sort(infarr, kind='quick'), np.sort(infarr, kind='heap'))

def test_sort_float16():
    arr = np.arange(65536, dtype=np.int16)
    temp = np.frombuffer(arr.tobytes(), dtype=np.float16)
    data = np.copy(temp)
    np.random.shuffle(data)
    data_backup = data
    assert_equal(np.sort(data, kind='quick'),
            np.sort(data_backup, kind='heap'))


@pytest.mark.parametrize("N", np.arange(1, 512))
@pytest.mark.parametrize("dtype", ['h', 'H', 'i', 'I', 'l', 'L'])
def test_sort_int(N, dtype):
    # Random data with MAX and MIN sprinkled
    minv = np.iinfo(dtype).min
    maxv = np.iinfo(dtype).max
    arr = np.random.randint(low=minv, high=maxv - 1, size=N, dtype=dtype)
    arr[np.random.choice(arr.shape[0], 10)] = minv
    arr[np.random.choice(arr.shape[0], 10)] = maxv
    assert_equal(np.sort(arr, kind='quick'), np.sort(arr, kind='heap'))


def test_sort_uint():
    # Random data with NPY_MAX_UINT32 sprinkled
    rng = np.random.default_rng(42)
    N = 2047
    maxv = np.iinfo(np.uint32).max
    arr = rng.integers(low=0, high=maxv, size=N).astype('uint32')
    arr[np.random.choice(arr.shape[0], 10)] = maxv
    assert_equal(np.sort(arr, kind='quick'), np.sort(arr, kind='heap'))

def test_private_get_ndarray_c_version():
    assert isinstance(_get_ndarray_c_version(), int)


@pytest.mark.parametrize("N", np.arange(1, 512))
@pytest.mark.parametrize("dtype", [np.float32, np.float64])
def test_argsort_float(N, dtype):
    rnd = np.random.RandomState(116112)
    # (1) Regular data with a few nan: doesn't use vectorized sort
    arr = -0.5 + rnd.random(N).astype(dtype)
    arr[rnd.choice(arr.shape[0], 3)] = np.nan
    assert_arg_sorted(arr, np.argsort(arr, kind='quick'))

    # (2) Random data with inf at the end of array
    # See: https://github.com/intel/x86-simd-sort/pull/39
    arr = -0.5 + rnd.rand(N).astype(dtype)
    arr[N - 1] = np.inf
    assert_arg_sorted(arr, np.argsort(arr, kind='quick'))


@pytest.mark.parametrize("N", np.arange(2, 512))
@pytest.mark.parametrize("dtype", [np.int32, np.uint32, np.int64, np.uint64])
def test_argsort_int(N, dtype):
    rnd = np.random.RandomState(1100710816)
    # (1) random data with min and max values
    minv = np.iinfo(dtype).min
    maxv = np.iinfo(dtype).max
    arr = rnd.randint(low=minv, high=maxv, size=N, dtype=dtype)
    i, j = rnd.choice(N, 2, replace=False)
    arr[i] = minv
    arr[j] = maxv
    assert_arg_sorted(arr, np.argsort(arr, kind='quick'))

    # (2) random data with max value at the end of array
    # See: https://github.com/intel/x86-simd-sort/pull/39
    arr = rnd.randint(low=minv, high=maxv, size=N, dtype=dtype)
    arr[N - 1] = maxv
    assert_arg_sorted(arr, np.argsort(arr, kind='quick'))

# Test large arrays that leverage openMP implementations from x86-simd-sort:
@pytest.mark.parametrize("dtype", [np.float16, np.float32, np.float64])
def test_sort_largearrays(dtype):
    N = 1000000
    rnd = np.random.RandomState(1100710816)
    arr = -0.5 + rnd.random(N).astype(dtype)
    assert_equal(np.sort(arr, kind='quick'), np.sort(arr, kind='heap'))

# Test large arrays that leverage openMP implementations from x86-simd-sort:
@pytest.mark.parametrize("dtype", [np.float32, np.float64])
def test_argsort_largearrays(dtype):
    N = 1000000
    rnd = np.random.RandomState(1100710816)
    arr = -0.5 + rnd.random(N).astype(dtype)
    assert_arg_sorted(arr, np.argsort(arr, kind='quick'))

@pytest.mark.skipif(not HAS_REFCOUNT, reason="Python lacks refcounts")
def test_gh_22683():
    b = 777.68760986
    a = np.array([b] * 10000, dtype=object)
    refc_start = sys.getrefcount(b)
    np.choose(np.zeros(10000, dtype=int), [a], out=a)
    np.choose(np.zeros(10000, dtype=int), [a], out=a)
    refc_end = sys.getrefcount(b)
    assert refc_end - refc_start < 10


def test_gh_24459():
    a = np.zeros((50, 3), dtype=np.float64)
    with pytest.raises(TypeError):
        np.choose(a, [3, -1])


def test_gh_28206():
    a = np.arange(3)
    b = np.ones((3, 3), dtype=np.int64)
    out = np.array([np.nan, np.nan, np.nan])

    with warnings.catch_warnings():
        warnings.simplefilter("error", RuntimeWarning)
        np.choose(a, b, out=out)


@pytest.mark.parametrize("N", np.arange(2, 512))
@pytest.mark.parametrize("dtype", [np.int16, np.uint16,
                        np.int32, np.uint32, np.int64, np.uint64])
def test_partition_int(N, dtype):
    rnd = np.random.RandomState(1100710816)
    # (1) random data with min and max values
    minv = np.iinfo(dtype).min
    maxv = np.iinfo(dtype).max
    arr = rnd.randint(low=minv, high=maxv, size=N, dtype=dtype)
    i, j = rnd.choice(N, 2, replace=False)
    arr[i] = minv
    arr[j] = maxv
    k = rnd.choice(N, 1)[0]
    assert_arr_partitioned(np.sort(arr)[k], k,
            np.partition(arr, k, kind='introselect'))
    assert_arr_partitioned(np.sort(arr)[k], k,
            arr[np.argpartition(arr, k, kind='introselect')])

    # (2) random data with max value at the end of array
    arr = rnd.randint(low=minv, high=maxv, size=N, dtype=dtype)
    arr[N - 1] = maxv
    assert_arr_partitioned(np.sort(arr)[k], k,
            np.partition(arr, k, kind='introselect'))
    assert_arr_partitioned(np.sort(arr)[k], k,
            arr[np.argpartition(arr, k, kind='introselect')])


@pytest.mark.parametrize("N", np.arange(2, 512))
@pytest.mark.parametrize("dtype", [np.float16, np.float32, np.float64])
def test_partition_fp(N, dtype):
    rnd = np.random.RandomState(1100710816)
    arr = -0.5 + rnd.random(N).astype(dtype)
    k = rnd.choice(N, 1)[0]
    assert_arr_partitioned(np.sort(arr)[k], k,
            np.partition(arr, k, kind='introselect'))
    assert_arr_partitioned(np.sort(arr)[k], k,
            arr[np.argpartition(arr, k, kind='introselect')])

    # Check that `np.inf < np.nan`
    # This follows np.sort
    arr[0] = np.nan
    arr[1] = np.inf
    o1 = np.partition(arr, -2, kind='introselect')
    o2 = arr[np.argpartition(arr, -2, kind='introselect')]
    for out in [o1, o2]:
        assert_(np.isnan(out[-1]))
        assert_equal(out[-2], np.inf)

def test_cannot_assign_data():
    a = np.arange(10)
    b = np.linspace(0, 1, 10)
    with pytest.raises(AttributeError):
        a.data = b.data

def test_insufficient_width():
    """
    If a 'width' parameter is passed into ``binary_repr`` that is insufficient
    to represent the number in base 2 (positive) or 2's complement (negative)
    form, the function used to silently ignore the parameter and return a
    representation using the minimal number of bits needed for the form in
    question. Such behavior is now considered unsafe from a user perspective
    and will raise an error.
    """
    with pytest.raises(ValueError):
        np.binary_repr(10, width=2)
    with pytest.raises(ValueError):
        np.binary_repr(-5, width=2)

def test_npy_char_raises():
    from numpy._core._multiarray_tests import npy_char_deprecation
    with pytest.raises(ValueError):
        npy_char_deprecation()


class TestDevice:
    """
    Test arr.device attribute and arr.to_device() method.
    """
    @pytest.mark.parametrize("func, arg", [
        (np.arange, 5),
        (np.empty_like, []),
        (np.zeros, 5),
        (np.empty, (5, 5)),
        (np.asarray, []),
        (np.asanyarray, []),
    ])
    def test_device(self, func, arg):
        arr = func(arg)
        assert arr.device == "cpu"
        arr = func(arg, device=None)
        assert arr.device == "cpu"
        arr = func(arg, device="cpu")
        assert arr.device == "cpu"

        with assert_raises_regex(
            ValueError,
            r"Device not understood. Only \"cpu\" is allowed, "
            r"but received: nonsense"
        ):
            func(arg, device="nonsense")

        with assert_raises_regex(
            AttributeError,
            r"attribute 'device' of '(numpy.|)ndarray' objects is "
            r"not writable"
        ):
            arr.device = "other"

    def test_to_device(self):
        arr = np.arange(5)

        assert arr.to_device("cpu") is arr
        with assert_raises_regex(
            ValueError,
            r"The stream argument in to_device\(\) is not supported"
        ):
            arr.to_device("cpu", stream=1)

def test_array_interface_excess_dimensions_raises():
    """Regression test for gh-27949: ensure too many dims raises ValueError instead of segfault."""

    # Dummy object to hold a custom __array_interface__
    class DummyArray:
        def __init__(self, interface):
            # Attach the array interface dict to mimic an array
            self.__array_interface__ = interface

    # Create a base array (scalar) and copy its interface
    base = np.array(42)  # base can be any scalar or array
    interface = dict(base.__array_interface__)

    # Modify the shape to exceed NumPy's dimension limit (NPY_MAXDIMS, typically 64)
    interface['shape'] = tuple([1] * 136)  # match the original bug report

    dummy = DummyArray(interface)
    # Now, using np.asanyarray on this dummy should trigger a ValueError (not segfault)
    with pytest.raises(ValueError, match="dimensions must be within"):
        np.asanyarray(dummy)

@pytest.mark.parametrize("dtype", [np.float32, np.float64, np.uint32, np.complex128])
def test_array_dunder_array_preserves_dtype_on_none(dtype):
    """
    Regression test for: https://github.com/numpy/numpy/issues/27407
    Ensure that __array__(None) returns an array of the same dtype.
    """
    a = np.array([1], dtype=dtype)
    b = a.__array__(None)
    assert_array_equal(a, b, strict=True)