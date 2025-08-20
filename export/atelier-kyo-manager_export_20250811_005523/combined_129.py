
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\concrete\guess.py ===
"""Various algorithms for helping identifying numbers and sequences."""


from sympy.concrete.products import (Product, product)
from sympy.core import Function, S
from sympy.core.add import Add
from sympy.core.numbers import Integer, Rational
from sympy.core.symbol import Symbol, symbols
from sympy.core.sympify import sympify
from sympy.functions.elementary.exponential import exp
from sympy.functions.elementary.integers import floor
from sympy.integrals.integrals import integrate
from sympy.polys.polyfuncs import rational_interpolate as rinterp
from sympy.polys.polytools import lcm
from sympy.simplify.radsimp import denom
from sympy.utilities import public


@public
def find_simple_recurrence_vector(l):
    """
    This function is used internally by other functions from the
    sympy.concrete.guess module. While most users may want to rather use the
    function find_simple_recurrence when looking for recurrence relations
    among rational numbers, the current function may still be useful when
    some post-processing has to be done.

    Explanation
    ===========

    The function returns a vector of length n when a recurrence relation of
    order n is detected in the sequence of rational numbers v.

    If the returned vector has a length 1, then the returned value is always
    the list [0], which means that no relation has been found.

    While the functions is intended to be used with rational numbers, it should
    work for other kinds of real numbers except for some cases involving
    quadratic numbers; for that reason it should be used with some caution when
    the argument is not a list of rational numbers.

    Examples
    ========

    >>> from sympy.concrete.guess import find_simple_recurrence_vector
    >>> from sympy import fibonacci
    >>> find_simple_recurrence_vector([fibonacci(k) for k in range(12)])
    [1, -1, -1]

    See Also
    ========

    See the function sympy.concrete.guess.find_simple_recurrence which is more
    user-friendly.

    """
    q1 = [0]
    q2 = [1]
    b, z = 0, len(l) >> 1
    while len(q2) <= z:
        while l[b]==0:
            b += 1
            if b == len(l):
                c = 1
                for x in q2:
                    c = lcm(c, denom(x))
                if q2[0]*c < 0: c = -c
                for k in range(len(q2)):
                    q2[k] = int(q2[k]*c)
                return q2
        a = S.One/l[b]
        m = [a]
        for k in range(b+1, len(l)):
            m.append(-sum(l[j+1]*m[b-j-1] for j in range(b, k))*a)
        l, m = m, [0] * max(len(q2), b+len(q1))
        for k, q in enumerate(q2):
            m[k] = a*q
        for k, q in enumerate(q1):
            m[k+b] += q
        while m[-1]==0: m.pop() # because trailing zeros can occur
        q1, q2, b = q2, m, 1
    return [0]

@public
def find_simple_recurrence(v, A=Function('a'), N=Symbol('n')):
    """
    Detects and returns a recurrence relation from a sequence of several integer
    (or rational) terms. The name of the function in the returned expression is
    'a' by default; the main variable is 'n' by default. The smallest index in
    the returned expression is always n (and never n-1, n-2, etc.).

    Examples
    ========

    >>> from sympy.concrete.guess import find_simple_recurrence
    >>> from sympy import fibonacci
    >>> find_simple_recurrence([fibonacci(k) for k in range(12)])
    -a(n) - a(n + 1) + a(n + 2)

    >>> from sympy import Function, Symbol
    >>> a = [1, 1, 1]
    >>> for k in range(15): a.append(5*a[-1]-3*a[-2]+8*a[-3])
    >>> find_simple_recurrence(a, A=Function('f'), N=Symbol('i'))
    -8*f(i) + 3*f(i + 1) - 5*f(i + 2) + f(i + 3)

    """
    p = find_simple_recurrence_vector(v)
    n = len(p)
    if n <= 1: return S.Zero

    return Add(*[A(N+n-1-k)*p[k] for k in range(n)])


@public
def rationalize(x, maxcoeff=10000):
    """
    Helps identifying a rational number from a float (or mpmath.mpf) value by
    using a continued fraction. The algorithm stops as soon as a large partial
    quotient is detected (greater than 10000 by default).

    Examples
    ========

    >>> from sympy.concrete.guess import rationalize
    >>> from mpmath import cos, pi
    >>> rationalize(cos(pi/3))
    1/2

    >>> from mpmath import mpf
    >>> rationalize(mpf("0.333333333333333"))
    1/3

    While the function is rather intended to help 'identifying' rational
    values, it may be used in some cases for approximating real numbers.
    (Though other functions may be more relevant in that case.)

    >>> rationalize(pi, maxcoeff = 250)
    355/113

    See Also
    ========

    Several other methods can approximate a real number as a rational, like:

      * fractions.Fraction.from_decimal
      * fractions.Fraction.from_float
      * mpmath.identify
      * mpmath.pslq by using the following syntax: mpmath.pslq([x, 1])
      * mpmath.findpoly by using the following syntax: mpmath.findpoly(x, 1)
      * sympy.simplify.nsimplify (which is a more general function)

    The main difference between the current function and all these variants is
    that control focuses on magnitude of partial quotients here rather than on
    global precision of the approximation. If the real is "known to be" a
    rational number, the current function should be able to detect it correctly
    with the default settings even when denominator is great (unless its
    expansion contains unusually big partial quotients) which may occur
    when studying sequences of increasing numbers. If the user cares more
    on getting simple fractions, other methods may be more convenient.

    """
    p0, p1 = 0, 1
    q0, q1 = 1, 0
    a = floor(x)
    while a < maxcoeff or q1==0:
        p = a*p1 + p0
        q = a*q1 + q0
        p0, p1 = p1, p
        q0, q1 = q1, q
        if x==a: break
        x = 1/(x-a)
        a = floor(x)
    return sympify(p) / q


@public
def guess_generating_function_rational(v, X=Symbol('x')):
    """
    Tries to "guess" a rational generating function for a sequence of rational
    numbers v.

    Examples
    ========

    >>> from sympy.concrete.guess import guess_generating_function_rational
    >>> from sympy import fibonacci
    >>> l = [fibonacci(k) for k in range(5,15)]
    >>> guess_generating_function_rational(l)
    (3*x + 5)/(-x**2 - x + 1)

    See Also
    ========

    sympy.series.approximants
    mpmath.pade

    """
    #   a) compute the denominator as q
    q = find_simple_recurrence_vector(v)
    n = len(q)
    if n <= 1: return None
    #   b) compute the numerator as p
    p = [sum(v[i-k]*q[k] for k in range(min(i+1, n)))
            for i in range(len(v)>>1)]
    return (sum(p[k]*X**k for k in range(len(p)))
            / sum(q[k]*X**k for k in range(n)))


@public
def guess_generating_function(v, X=Symbol('x'), types=['all'], maxsqrtn=2):
    """
    Tries to "guess" a generating function for a sequence of rational numbers v.
    Only a few patterns are implemented yet.

    Explanation
    ===========

    The function returns a dictionary where keys are the name of a given type of
    generating function. Six types are currently implemented:

         type  |  formal definition
        -------+----------------------------------------------------------------
        ogf    | f(x) = Sum(            a_k * x^k       ,  k: 0..infinity )
        egf    | f(x) = Sum(            a_k * x^k / k!  ,  k: 0..infinity )
        lgf    | f(x) = Sum( (-1)^(k+1) a_k * x^k / k   ,  k: 1..infinity )
               |        (with initial index being hold as 1 rather than 0)
        hlgf   | f(x) = Sum(            a_k * x^k / k   ,  k: 1..infinity )
               |        (with initial index being hold as 1 rather than 0)
        lgdogf | f(x) = derivate( log(Sum( a_k * x^k, k: 0..infinity )), x)
        lgdegf | f(x) = derivate( log(Sum( a_k * x^k / k!, k: 0..infinity )), x)

    In order to spare time, the user can select only some types of generating
    functions (default being ['all']). While forgetting to use a list in the
    case of a single type may seem to work most of the time as in: types='ogf'
    this (convenient) syntax may lead to unexpected extra results in some cases.

    Discarding a type when calling the function does not mean that the type will
    not be present in the returned dictionary; it only means that no extra
    computation will be performed for that type, but the function may still add
    it in the result when it can be easily converted from another type.

    Two generating functions (lgdogf and lgdegf) are not even computed if the
    initial term of the sequence is 0; it may be useful in that case to try
    again after having removed the leading zeros.

    Examples
    ========

    >>> from sympy.concrete.guess import guess_generating_function as ggf
    >>> ggf([k+1 for k in range(12)], types=['ogf', 'lgf', 'hlgf'])
    {'hlgf': 1/(1 - x), 'lgf': 1/(x + 1), 'ogf': 1/(x**2 - 2*x + 1)}

    >>> from sympy import sympify
    >>> l = sympify("[3/2, 11/2, 0, -121/2, -363/2, 121]")
    >>> ggf(l)
    {'ogf': (x + 3/2)/(11*x**2 - 3*x + 1)}

    >>> from sympy import fibonacci
    >>> ggf([fibonacci(k) for k in range(5, 15)], types=['ogf'])
    {'ogf': (3*x + 5)/(-x**2 - x + 1)}

    >>> from sympy import factorial
    >>> ggf([factorial(k) for k in range(12)], types=['ogf', 'egf', 'lgf'])
    {'egf': 1/(1 - x)}

    >>> ggf([k+1 for k in range(12)], types=['egf'])
    {'egf': (x + 1)*exp(x), 'lgdegf': (x + 2)/(x + 1)}

    N-th root of a rational function can also be detected (below is an example
    coming from the sequence A108626 from https://oeis.org).
    The greatest n-th root to be tested is specified as maxsqrtn (default 2).

    >>> ggf([1, 2, 5, 14, 41, 124, 383, 1200, 3799, 12122, 38919])['ogf']
    sqrt(1/(x**4 + 2*x**2 - 4*x + 1))

    References
    ==========

    .. [1] "Concrete Mathematics", R.L. Graham, D.E. Knuth, O. Patashnik
    .. [2] https://oeis.org/wiki/Generating_functions

    """
    # List of all types of all g.f. known by the algorithm
    if 'all' in types:
        types = ('ogf', 'egf', 'lgf', 'hlgf', 'lgdogf', 'lgdegf')

    result = {}

    # Ordinary Generating Function (ogf)
    if 'ogf' in types:
        # Perform some convolutions of the sequence with itself
        t = [1] + [0]*(len(v) - 1)
        for d in range(max(1, maxsqrtn)):
            t = [sum(t[n-i]*v[i] for i in range(n+1)) for n in range(len(v))]
            g = guess_generating_function_rational(t, X=X)
            if g:
                result['ogf'] = g**Rational(1, d+1)
                break

    # Exponential Generating Function (egf)
    if 'egf' in types:
        # Transform sequence (division by factorial)
        w, f = [], S.One
        for i, k in enumerate(v):
            f *= i if i else 1
            w.append(k/f)
        # Perform some convolutions of the sequence with itself
        t = [1] + [0]*(len(w) - 1)
        for d in range(max(1, maxsqrtn)):
            t = [sum(t[n-i]*w[i] for i in range(n+1)) for n in range(len(w))]
            g = guess_generating_function_rational(t, X=X)
            if g:
                result['egf'] = g**Rational(1, d+1)
                break

    # Logarithmic Generating Function (lgf)
    if 'lgf' in types:
        # Transform sequence (multiplication by (-1)^(n+1) / n)
        w, f = [], S.NegativeOne
        for i, k in enumerate(v):
            f = -f
            w.append(f*k/Integer(i+1))
        # Perform some convolutions of the sequence with itself
        t = [1] + [0]*(len(w) - 1)
        for d in range(max(1, maxsqrtn)):
            t = [sum(t[n-i]*w[i] for i in range(n+1)) for n in range(len(w))]
            g = guess_generating_function_rational(t, X=X)
            if g:
                result['lgf'] = g**Rational(1, d+1)
                break

    # Hyperbolic logarithmic Generating Function (hlgf)
    if 'hlgf' in types:
        # Transform sequence (division by n+1)
        w = []
        for i, k in enumerate(v):
            w.append(k/Integer(i+1))
        # Perform some convolutions of the sequence with itself
        t = [1] + [0]*(len(w) - 1)
        for d in range(max(1, maxsqrtn)):
            t = [sum(t[n-i]*w[i] for i in range(n+1)) for n in range(len(w))]
            g = guess_generating_function_rational(t, X=X)
            if g:
                result['hlgf'] = g**Rational(1, d+1)
                break

    # Logarithmic derivative of ordinary generating Function (lgdogf)
    if v[0] != 0 and ('lgdogf' in types
                       or ('ogf' in types and 'ogf' not in result)):
        # Transform sequence by computing f'(x)/f(x)
        # because log(f(x)) = integrate( f'(x)/f(x) )
        a, w = sympify(v[0]), []
        for n in range(len(v)-1):
            w.append(
               (v[n+1]*(n+1) - sum(w[-i-1]*v[i+1] for i in range(n)))/a)
        # Perform some convolutions of the sequence with itself
        t = [1] + [0]*(len(w) - 1)
        for d in range(max(1, maxsqrtn)):
            t = [sum(t[n-i]*w[i] for i in range(n+1)) for n in range(len(w))]
            g = guess_generating_function_rational(t, X=X)
            if g:
                result['lgdogf'] = g**Rational(1, d+1)
                if 'ogf' not in result:
                    result['ogf'] = exp(integrate(result['lgdogf'], X))
                break

    # Logarithmic derivative of exponential generating Function (lgdegf)
    if v[0] != 0 and ('lgdegf' in types
                       or ('egf' in types and 'egf' not in result)):
        # Transform sequence / step 1 (division by factorial)
        z, f = [], S.One
        for i, k in enumerate(v):
            f *= i if i else 1
            z.append(k/f)
        # Transform sequence / step 2 by computing f'(x)/f(x)
        # because log(f(x)) = integrate( f'(x)/f(x) )
        a, w = z[0], []
        for n in range(len(z)-1):
            w.append(
               (z[n+1]*(n+1) - sum(w[-i-1]*z[i+1] for i in range(n)))/a)
        # Perform some convolutions of the sequence with itself
        t = [1] + [0]*(len(w) - 1)
        for d in range(max(1, maxsqrtn)):
            t = [sum(t[n-i]*w[i] for i in range(n+1)) for n in range(len(w))]
            g = guess_generating_function_rational(t, X=X)
            if g:
                result['lgdegf'] = g**Rational(1, d+1)
                if 'egf' not in result:
                    result['egf'] = exp(integrate(result['lgdegf'], X))
                break

    return result


@public
def guess(l, all=False, evaluate=True, niter=2, variables=None):
    """
    This function is adapted from the Rate.m package for Mathematica
    written by Christian Krattenthaler.
    It tries to guess a formula from a given sequence of rational numbers.

    Explanation
    ===========

    In order to speed up the process, the 'all' variable is set to False by
    default, stopping the computation as some results are returned during an
    iteration; the variable can be set to True if more iterations are needed
    (other formulas may be found; however they may be equivalent to the first
    ones).

    Another option is the 'evaluate' variable (default is True); setting it
    to False will leave the involved products unevaluated.

    By default, the number of iterations is set to 2 but a greater value (up
    to len(l)-1) can be specified with the optional 'niter' variable.
    More and more convoluted results are found when the order of the
    iteration gets higher:

      * first iteration returns polynomial or rational functions;
      * second iteration returns products of rising factorials and their
        inverses;
      * third iteration returns products of products of rising factorials
        and their inverses;
      * etc.

    The returned formulas contain symbols i0, i1, i2, ... where the main
    variables is i0 (and auxiliary variables are i1, i2, ...). A list of
    other symbols can be provided in the 'variables' option; the length of
    the least should be the value of 'niter' (more is acceptable but only
    the first symbols will be used); in this case, the main variable will be
    the first symbol in the list.

    Examples
    ========

    >>> from sympy.concrete.guess import guess
    >>> guess([1,2,6,24,120], evaluate=False)
    [Product(i1 + 1, (i1, 1, i0 - 1))]

    >>> from sympy import symbols
    >>> r = guess([1,2,7,42,429,7436,218348,10850216], niter=4)
    >>> i0 = symbols("i0")
    >>> [r[0].subs(i0,n).doit() for n in range(1,10)]
    [1, 2, 7, 42, 429, 7436, 218348, 10850216, 911835460]
    """
    if any(a==0 for a in l[:-1]):
        return []
    N = len(l)
    niter = min(N-1, niter)
    myprod = product if evaluate else Product
    g = []
    res = []
    if variables is None:
        symb = symbols('i:'+str(niter))
    else:
        symb = variables
    for k, s in enumerate(symb):
        g.append(l)
        n, r = len(l), []
        for i in range(n-2-1, -1, -1):
            ri = rinterp(enumerate(g[k][:-1], start=1), i, X=s)
            if ((denom(ri).subs({s:n}) != 0)
                    and (ri.subs({s:n}) - g[k][-1] == 0)
                    and ri not in r):
                r.append(ri)
        if r:
            for i in range(k-1, -1, -1):
                r = [g[i][0]
                      * myprod(v, (symb[i+1], 1, symb[i]-1)) for v in r]
            if not all: return r
            res += r
        l = [Rational(l[i+1], l[i]) for i in range(N-k-1)]
    return res

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\sparse.py ===
from collections.abc import Callable

from sympy.core.containers import Dict
from sympy.utilities.exceptions import sympy_deprecation_warning
from sympy.utilities.iterables import is_sequence
from sympy.utilities.misc import as_int

from .matrixbase import MatrixBase
from .repmatrix import MutableRepMatrix, RepMatrix

from .utilities import _iszero

from .decompositions import (
    _liupc, _row_structure_symbolic_cholesky, _cholesky_sparse,
    _LDLdecomposition_sparse)

from .solvers import (
    _lower_triangular_solve_sparse, _upper_triangular_solve_sparse)


class SparseRepMatrix(RepMatrix):
    """
    A sparse matrix (a matrix with a large number of zero elements).

    Examples
    ========

    >>> from sympy import SparseMatrix, ones
    >>> SparseMatrix(2, 2, range(4))
    Matrix([
    [0, 1],
    [2, 3]])
    >>> SparseMatrix(2, 2, {(1, 1): 2})
    Matrix([
    [0, 0],
    [0, 2]])

    A SparseMatrix can be instantiated from a ragged list of lists:

    >>> SparseMatrix([[1, 2, 3], [1, 2], [1]])
    Matrix([
    [1, 2, 3],
    [1, 2, 0],
    [1, 0, 0]])

    For safety, one may include the expected size and then an error
    will be raised if the indices of any element are out of range or
    (for a flat list) if the total number of elements does not match
    the expected shape:

    >>> SparseMatrix(2, 2, [1, 2])
    Traceback (most recent call last):
    ...
    ValueError: List length (2) != rows*columns (4)

    Here, an error is not raised because the list is not flat and no
    element is out of range:

    >>> SparseMatrix(2, 2, [[1, 2]])
    Matrix([
    [1, 2],
    [0, 0]])

    But adding another element to the first (and only) row will cause
    an error to be raised:

    >>> SparseMatrix(2, 2, [[1, 2, 3]])
    Traceback (most recent call last):
    ...
    ValueError: The location (0, 2) is out of designated range: (1, 1)

    To autosize the matrix, pass None for rows:

    >>> SparseMatrix(None, [[1, 2, 3]])
    Matrix([[1, 2, 3]])
    >>> SparseMatrix(None, {(1, 1): 1, (3, 3): 3})
    Matrix([
    [0, 0, 0, 0],
    [0, 1, 0, 0],
    [0, 0, 0, 0],
    [0, 0, 0, 3]])

    Values that are themselves a Matrix are automatically expanded:

    >>> SparseMatrix(4, 4, {(1, 1): ones(2)})
    Matrix([
    [0, 0, 0, 0],
    [0, 1, 1, 0],
    [0, 1, 1, 0],
    [0, 0, 0, 0]])

    A ValueError is raised if the expanding matrix tries to overwrite
    a different element already present:

    >>> SparseMatrix(3, 3, {(0, 0): ones(2), (1, 1): 2})
    Traceback (most recent call last):
    ...
    ValueError: collision at (1, 1)

    See Also
    ========
    DenseMatrix
    MutableSparseMatrix
    ImmutableSparseMatrix
    """

    @classmethod
    def _handle_creation_inputs(cls, *args, **kwargs):
        if len(args) == 1 and isinstance(args[0], MatrixBase):
            rows = args[0].rows
            cols = args[0].cols
            smat = args[0].todok()
            return rows, cols, smat

        smat = {}
        # autosizing
        if len(args) == 2 and args[0] is None:
            args = [None, None, args[1]]

        if len(args) == 3:
            r, c = args[:2]
            if r is c is None:
                rows = cols = None
            elif None in (r, c):
                raise ValueError(
                    'Pass rows=None and no cols for autosizing.')
            else:
                rows, cols = as_int(args[0]), as_int(args[1])

            if isinstance(args[2], Callable):
                op = args[2]

                if None in (rows, cols):
                    raise ValueError(
                        "{} and {} must be integers for this "
                        "specification.".format(rows, cols))

                row_indices = [cls._sympify(i) for i in range(rows)]
                col_indices = [cls._sympify(j) for j in range(cols)]

                for i in row_indices:
                    for j in col_indices:
                        value = cls._sympify(op(i, j))
                        if value != cls.zero:
                            smat[i, j] = value

                return rows, cols, smat

            elif isinstance(args[2], (dict, Dict)):
                def update(i, j, v):
                    # update smat and make sure there are no collisions
                    if v:
                        if (i, j) in smat and v != smat[i, j]:
                            raise ValueError(
                                "There is a collision at {} for {} and {}."
                                .format((i, j), v, smat[i, j])
                            )
                        smat[i, j] = v

                # manual copy, copy.deepcopy() doesn't work
                for (r, c), v in args[2].items():
                    if isinstance(v, MatrixBase):
                        for (i, j), vv in v.todok().items():
                            update(r + i, c + j, vv)
                    elif isinstance(v, (list, tuple)):
                        _, _, smat = cls._handle_creation_inputs(v, **kwargs)
                        for i, j in smat:
                            update(r + i, c + j, smat[i, j])
                    else:
                        v = cls._sympify(v)
                        update(r, c, cls._sympify(v))

            elif is_sequence(args[2]):
                flat = not any(is_sequence(i) for i in args[2])
                if not flat:
                    _, _, smat = \
                        cls._handle_creation_inputs(args[2], **kwargs)
                else:
                    flat_list = args[2]
                    if len(flat_list) != rows * cols:
                        raise ValueError(
                            "The length of the flat list ({}) does not "
                            "match the specified size ({} * {})."
                            .format(len(flat_list), rows, cols)
                        )

                    for i in range(rows):
                        for j in range(cols):
                            value = flat_list[i*cols + j]
                            value = cls._sympify(value)
                            if value != cls.zero:
                                smat[i, j] = value

            if rows is None:  # autosizing
                keys = smat.keys()
                rows = max(r for r, _ in keys) + 1 if keys else 0
                cols = max(c for _, c in keys) + 1 if keys else 0

            else:
                for i, j in smat.keys():
                    if i and i >= rows or j and j >= cols:
                        raise ValueError(
                            "The location {} is out of the designated range"
                            "[{}, {}]x[{}, {}]"
                            .format((i, j), 0, rows - 1, 0, cols - 1)
                        )

            return rows, cols, smat

        elif len(args) == 1 and isinstance(args[0], (list, tuple)):
            # list of values or lists
            v = args[0]
            c = 0
            for i, row in enumerate(v):
                if not isinstance(row, (list, tuple)):
                    row = [row]
                for j, vv in enumerate(row):
                    if vv != cls.zero:
                        smat[i, j] = cls._sympify(vv)
                c = max(c, len(row))
            rows = len(v) if c else 0
            cols = c
            return rows, cols, smat

        else:
            # handle full matrix forms with _handle_creation_inputs
            rows, cols, mat = super()._handle_creation_inputs(*args)
            for i in range(rows):
                for j in range(cols):
                    value = mat[cols*i + j]
                    if value != cls.zero:
                        smat[i, j] = value

            return rows, cols, smat

    @property
    def _smat(self):

        sympy_deprecation_warning(
            """
            The private _smat attribute of SparseMatrix is deprecated. Use the
            .todok() method instead.
            """,
            deprecated_since_version="1.9",
            active_deprecations_target="deprecated-private-matrix-attributes"
        )

        return self.todok()

    def _eval_inverse(self, **kwargs):
        return self.inv(method=kwargs.get('method', 'LDL'),
                        iszerofunc=kwargs.get('iszerofunc', _iszero),
                        try_block_diag=kwargs.get('try_block_diag', False))

    def applyfunc(self, f):
        """Apply a function to each element of the matrix.

        Examples
        ========

        >>> from sympy import SparseMatrix
        >>> m = SparseMatrix(2, 2, lambda i, j: i*2+j)
        >>> m
        Matrix([
        [0, 1],
        [2, 3]])
        >>> m.applyfunc(lambda i: 2*i)
        Matrix([
        [0, 2],
        [4, 6]])

        """
        if not callable(f):
            raise TypeError("`f` must be callable.")

        # XXX: This only applies the function to the nonzero elements of the
        # matrix so is inconsistent with DenseMatrix.applyfunc e.g.
        #   zeros(2, 2).applyfunc(lambda x: x + 1)
        dok = {}
        for k, v in self.todok().items():
            fv = f(v)
            if fv != 0:
                dok[k] = fv

        return self._new(self.rows, self.cols, dok)

    def as_immutable(self):
        """Returns an Immutable version of this Matrix."""
        from .immutable import ImmutableSparseMatrix
        return ImmutableSparseMatrix(self)

    def as_mutable(self):
        """Returns a mutable version of this matrix.

        Examples
        ========

        >>> from sympy import ImmutableMatrix
        >>> X = ImmutableMatrix([[1, 2], [3, 4]])
        >>> Y = X.as_mutable()
        >>> Y[1, 1] = 5 # Can set values in Y
        >>> Y
        Matrix([
        [1, 2],
        [3, 5]])
        """
        return MutableSparseMatrix(self)

    def col_list(self):
        """Returns a column-sorted list of non-zero elements of the matrix.

        Examples
        ========

        >>> from sympy import SparseMatrix
        >>> a=SparseMatrix(((1, 2), (3, 4)))
        >>> a
        Matrix([
        [1, 2],
        [3, 4]])
        >>> a.CL
        [(0, 0, 1), (1, 0, 3), (0, 1, 2), (1, 1, 4)]

        See Also
        ========

        sympy.matrices.sparse.SparseMatrix.row_list
        """
        return [tuple(k + (self[k],)) for k in sorted(self.todok().keys(), key=lambda k: list(reversed(k)))]

    def nnz(self):
        """Returns the number of non-zero elements in Matrix."""
        return len(self.todok())

    def row_list(self):
        """Returns a row-sorted list of non-zero elements of the matrix.

        Examples
        ========

        >>> from sympy import SparseMatrix
        >>> a = SparseMatrix(((1, 2), (3, 4)))
        >>> a
        Matrix([
        [1, 2],
        [3, 4]])
        >>> a.RL
        [(0, 0, 1), (0, 1, 2), (1, 0, 3), (1, 1, 4)]

        See Also
        ========

        sympy.matrices.sparse.SparseMatrix.col_list
        """
        return [tuple(k + (self[k],)) for k in
            sorted(self.todok().keys(), key=list)]

    def scalar_multiply(self, scalar):
        "Scalar element-wise multiplication"
        return scalar * self

    def solve_least_squares(self, rhs, method='LDL'):
        """Return the least-square fit to the data.

        By default the cholesky_solve routine is used (method='CH'); other
        methods of matrix inversion can be used. To find out which are
        available, see the docstring of the .inv() method.

        Examples
        ========

        >>> from sympy import SparseMatrix, Matrix, ones
        >>> A = Matrix([1, 2, 3])
        >>> B = Matrix([2, 3, 4])
        >>> S = SparseMatrix(A.row_join(B))
        >>> S
        Matrix([
        [1, 2],
        [2, 3],
        [3, 4]])

        If each line of S represent coefficients of Ax + By
        and x and y are [2, 3] then S*xy is:

        >>> r = S*Matrix([2, 3]); r
        Matrix([
        [ 8],
        [13],
        [18]])

        But let's add 1 to the middle value and then solve for the
        least-squares value of xy:

        >>> xy = S.solve_least_squares(Matrix([8, 14, 18])); xy
        Matrix([
        [ 5/3],
        [10/3]])

        The error is given by S*xy - r:

        >>> S*xy - r
        Matrix([
        [1/3],
        [1/3],
        [1/3]])
        >>> _.norm().n(2)
        0.58

        If a different xy is used, the norm will be higher:

        >>> xy += ones(2, 1)/10
        >>> (S*xy - r).norm().n(2)
        1.5

        """
        t = self.T
        return (t*self).inv(method=method)*t*rhs

    def solve(self, rhs, method='LDL'):
        """Return solution to self*soln = rhs using given inversion method.

        For a list of possible inversion methods, see the .inv() docstring.
        """
        if not self.is_square:
            if self.rows < self.cols:
                raise ValueError('Under-determined system.')
            elif self.rows > self.cols:
                raise ValueError('For over-determined system, M, having '
                    'more rows than columns, try M.solve_least_squares(rhs).')
        else:
            return self.inv(method=method).multiply(rhs)

    RL = property(row_list, None, None, "Alternate faster representation")
    CL = property(col_list, None, None, "Alternate faster representation")

    def liupc(self):
        return _liupc(self)

    def row_structure_symbolic_cholesky(self):
        return _row_structure_symbolic_cholesky(self)

    def cholesky(self, hermitian=True):
        return _cholesky_sparse(self, hermitian=hermitian)

    def LDLdecomposition(self, hermitian=True):
        return _LDLdecomposition_sparse(self, hermitian=hermitian)

    def lower_triangular_solve(self, rhs):
        return _lower_triangular_solve_sparse(self, rhs)

    def upper_triangular_solve(self, rhs):
        return _upper_triangular_solve_sparse(self, rhs)

    liupc.__doc__                           = _liupc.__doc__
    row_structure_symbolic_cholesky.__doc__ = _row_structure_symbolic_cholesky.__doc__
    cholesky.__doc__                        = _cholesky_sparse.__doc__
    LDLdecomposition.__doc__                = _LDLdecomposition_sparse.__doc__
    lower_triangular_solve.__doc__          = lower_triangular_solve.__doc__
    upper_triangular_solve.__doc__          = upper_triangular_solve.__doc__


class MutableSparseMatrix(SparseRepMatrix, MutableRepMatrix):

    @classmethod
    def _new(cls, *args, **kwargs):
        rows, cols, smat = cls._handle_creation_inputs(*args, **kwargs)

        rep = cls._smat_to_DomainMatrix(rows, cols, smat)

        return cls._fromrep(rep)


SparseMatrix = MutableSparseMatrix

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\multivariate_resultants.py ===
"""
This module contains functions for two multivariate resultants. These
are:

- Dixon's resultant.
- Macaulay's resultant.

Multivariate resultants are used to identify whether a multivariate
system has common roots. That is when the resultant is equal to zero.
"""
from math import prod

from sympy.core.mul import Mul
from sympy.matrices.dense import (Matrix, diag)
from sympy.polys.polytools import (Poly, degree_list, rem)
from sympy.simplify.simplify import simplify
from sympy.tensor.indexed import IndexedBase
from sympy.polys.monomials import itermonomials, monomial_deg
from sympy.polys.orderings import monomial_key
from sympy.polys.polytools import poly_from_expr, total_degree
from sympy.functions.combinatorial.factorials import binomial
from itertools import combinations_with_replacement
from sympy.utilities.exceptions import sympy_deprecation_warning

class DixonResultant():
    """
    A class for retrieving the Dixon's resultant of a multivariate
    system.

    Examples
    ========

    >>> from sympy import symbols

    >>> from sympy.polys.multivariate_resultants import DixonResultant
    >>> x, y = symbols('x, y')

    >>> p = x + y
    >>> q = x ** 2 + y ** 3
    >>> h = x ** 2 + y

    >>> dixon = DixonResultant(variables=[x, y], polynomials=[p, q, h])
    >>> poly = dixon.get_dixon_polynomial()
    >>> matrix = dixon.get_dixon_matrix(polynomial=poly)
    >>> matrix
    Matrix([
    [ 0,  0, -1,  0, -1],
    [ 0, -1,  0, -1,  0],
    [-1,  0,  1,  0,  0],
    [ 0, -1,  0,  0,  1],
    [-1,  0,  0,  1,  0]])
    >>> matrix.det()
    0

    See Also
    ========

    Notebook in examples: sympy/example/notebooks.

    References
    ==========

    .. [1] [Kapur1994]_
    .. [2] [Palancz08]_

    """

    def __init__(self, polynomials, variables):
        """
        A class that takes two lists, a list of polynomials and list of
        variables. Returns the Dixon matrix of the multivariate system.

        Parameters
        ----------
        polynomials : list of polynomials
            A  list of m n-degree polynomials
        variables: list
            A list of all n variables
        """
        self.polynomials = polynomials
        self.variables = variables

        self.n = len(self.variables)
        self.m = len(self.polynomials)

        a = IndexedBase("alpha")
        # A list of n alpha variables (the replacing variables)
        self.dummy_variables = [a[i] for i in range(self.n)]

        # A list of the d_max of each variable.
        self._max_degrees = [max(degree_list(poly)[i] for poly in self.polynomials)
            for i in range(self.n)]

    @property
    def max_degrees(self):
        sympy_deprecation_warning(
            """
            The max_degrees property of DixonResultant is deprecated.
            """,
            deprecated_since_version="1.5",
            active_deprecations_target="deprecated-dixonresultant-properties",
        )
        return self._max_degrees

    def get_dixon_polynomial(self):
        r"""
        Returns
        =======

        dixon_polynomial: polynomial
            Dixon's polynomial is calculated as:

            delta = Delta(A) / ((x_1 - a_1) ... (x_n - a_n)) where,

            A =  |p_1(x_1,... x_n), ..., p_n(x_1,... x_n)|
                 |p_1(a_1,... x_n), ..., p_n(a_1,... x_n)|
                 |...             , ...,              ...|
                 |p_1(a_1,... a_n), ..., p_n(a_1,... a_n)|
        """
        if self.m != (self.n + 1):
            raise ValueError('Method invalid for given combination.')

        # First row
        rows = [self.polynomials]

        temp = list(self.variables)

        for idx in range(self.n):
            temp[idx] = self.dummy_variables[idx]
            substitution = dict(zip(self.variables, temp))
            rows.append([f.subs(substitution) for f in self.polynomials])

        A = Matrix(rows)

        terms = zip(self.variables, self.dummy_variables)
        product_of_differences = Mul(*[a - b for a, b in terms])
        dixon_polynomial = (A.det() / product_of_differences).factor()

        return poly_from_expr(dixon_polynomial, self.dummy_variables)[0]

    def get_upper_degree(self):
        sympy_deprecation_warning(
            """
            The get_upper_degree() method of DixonResultant is deprecated. Use
            get_max_degrees() instead.
            """,
            deprecated_since_version="1.5",
            active_deprecations_target="deprecated-dixonresultant-properties"
        )
        list_of_products = [self.variables[i] ** self._max_degrees[i]
                            for i in range(self.n)]
        product = prod(list_of_products)
        product = Poly(product).monoms()

        return monomial_deg(*product)

    def get_max_degrees(self, polynomial):
        r"""
        Returns a list of the maximum degree of each variable appearing
        in the coefficients of the Dixon polynomial. The coefficients are
        viewed as polys in $x_1, x_2, \dots, x_n$.
        """
        deg_lists = [degree_list(Poly(poly, self.variables))
                     for poly in polynomial.coeffs()]

        max_degrees = [max(degs) for degs in zip(*deg_lists)]

        return max_degrees

    def get_dixon_matrix(self, polynomial):
        r"""
        Construct the Dixon matrix from the coefficients of polynomial
        \alpha. Each coefficient is viewed as a polynomial of x_1, ...,
        x_n.
        """

        max_degrees = self.get_max_degrees(polynomial)

        # list of column headers of the Dixon matrix.
        monomials = itermonomials(self.variables, max_degrees)
        monomials = sorted(monomials, reverse=True,
                           key=monomial_key('lex', self.variables))

        dixon_matrix = Matrix([[Poly(c, *self.variables).coeff_monomial(m)
                                for m in monomials]
                                for c in polynomial.coeffs()])

        # remove columns if needed
        if dixon_matrix.shape[0] != dixon_matrix.shape[1]:
            keep = [column for column in range(dixon_matrix.shape[-1])
                    if any(element != 0 for element
                        in dixon_matrix[:, column])]

            dixon_matrix = dixon_matrix[:, keep]

        return dixon_matrix

    def KSY_precondition(self, matrix):
        """
        Test for the validity of the Kapur-Saxena-Yang precondition.

        The precondition requires that the column corresponding to the
        monomial 1 = x_1 ^ 0 * x_2 ^ 0 * ... * x_n ^ 0 is not a linear
        combination of the remaining ones. In SymPy notation this is
        the last column. For the precondition to hold the last non-zero
        row of the rref matrix should be of the form [0, 0, ..., 1].
        """
        if matrix.is_zero_matrix:
            return False

        m, n = matrix.shape

        # simplify the matrix and keep only its non-zero rows
        matrix = simplify(matrix.rref()[0])
        rows = [i for i in range(m) if any(matrix[i, j] != 0 for j in range(n))]
        matrix = matrix[rows,:]

        condition = Matrix([[0]*(n-1) + [1]])

        if matrix[-1,:] == condition:
            return True
        else:
            return False

    def delete_zero_rows_and_columns(self, matrix):
        """Remove the zero rows and columns of the matrix."""
        rows = [
            i for i in range(matrix.rows) if not matrix.row(i).is_zero_matrix]
        cols = [
            j for j in range(matrix.cols) if not matrix.col(j).is_zero_matrix]

        return matrix[rows, cols]

    def product_leading_entries(self, matrix):
        """Calculate the product of the leading entries of the matrix."""
        res = 1
        for row in range(matrix.rows):
            for el in matrix.row(row):
                if el != 0:
                    res = res * el
                    break
        return res

    def get_KSY_Dixon_resultant(self, matrix):
        """Calculate the Kapur-Saxena-Yang approach to the Dixon Resultant."""
        matrix = self.delete_zero_rows_and_columns(matrix)
        _, U, _ = matrix.LUdecomposition()
        matrix = self.delete_zero_rows_and_columns(simplify(U))

        return self.product_leading_entries(matrix)

class MacaulayResultant():
    """
    A class for calculating the Macaulay resultant. Note that the
    polynomials must be homogenized and their coefficients must be
    given as symbols.

    Examples
    ========

    >>> from sympy import symbols

    >>> from sympy.polys.multivariate_resultants import MacaulayResultant
    >>> x, y, z = symbols('x, y, z')

    >>> a_0, a_1, a_2 = symbols('a_0, a_1, a_2')
    >>> b_0, b_1, b_2 = symbols('b_0, b_1, b_2')
    >>> c_0, c_1, c_2,c_3, c_4 = symbols('c_0, c_1, c_2, c_3, c_4')

    >>> f = a_0 * y -  a_1 * x + a_2 * z
    >>> g = b_1 * x ** 2 + b_0 * y ** 2 - b_2 * z ** 2
    >>> h = c_0 * y * z ** 2 - c_1 * x ** 3 + c_2 * x ** 2 * z - c_3 * x * z ** 2 + c_4 * z ** 3

    >>> mac = MacaulayResultant(polynomials=[f, g, h], variables=[x, y, z])
    >>> mac.monomial_set
    [x**4, x**3*y, x**3*z, x**2*y**2, x**2*y*z, x**2*z**2, x*y**3,
    x*y**2*z, x*y*z**2, x*z**3, y**4, y**3*z, y**2*z**2, y*z**3, z**4]
    >>> matrix = mac.get_matrix()
    >>> submatrix = mac.get_submatrix(matrix)
    >>> submatrix
    Matrix([
    [-a_1,  a_0,  a_2,    0],
    [   0, -a_1,    0,    0],
    [   0,    0, -a_1,    0],
    [   0,    0,    0, -a_1]])

    See Also
    ========

    Notebook in examples: sympy/example/notebooks.

    References
    ==========

    .. [1] [Bruce97]_
    .. [2] [Stiller96]_

    """
    def __init__(self, polynomials, variables):
        """
        Parameters
        ==========

        variables: list
            A list of all n variables
        polynomials : list of SymPy polynomials
            A  list of m n-degree polynomials
        """
        self.polynomials = polynomials
        self.variables = variables
        self.n = len(variables)

        # A list of the d_max of each polynomial.
        self.degrees = [total_degree(poly, *self.variables) for poly
                        in self.polynomials]

        self.degree_m = self._get_degree_m()
        self.monomials_size = self.get_size()

        # The set T of all possible monomials of degree degree_m
        self.monomial_set = self.get_monomials_of_certain_degree(self.degree_m)

    def _get_degree_m(self):
        r"""
        Returns
        =======

        degree_m: int
            The degree_m is calculated as  1 + \sum_1 ^ n (d_i - 1),
            where d_i is the degree of the i polynomial
        """
        return 1 + sum(d - 1 for d in self.degrees)

    def get_size(self):
        r"""
        Returns
        =======

        size: int
            The size of set T. Set T is the set of all possible
            monomials of the n variables for degree equal to the
            degree_m
        """
        return binomial(self.degree_m + self.n - 1, self.n - 1)

    def get_monomials_of_certain_degree(self, degree):
        """
        Returns
        =======

        monomials: list
            A list of monomials of a certain degree.
        """
        monomials = [Mul(*monomial) for monomial
                     in combinations_with_replacement(self.variables,
                                                      degree)]

        return sorted(monomials, reverse=True,
                      key=monomial_key('lex', self.variables))

    def get_row_coefficients(self):
        """
        Returns
        =======

        row_coefficients: list
            The row coefficients of Macaulay's matrix
        """
        row_coefficients = []
        divisible = []
        for i in range(self.n):
            if i == 0:
                degree = self.degree_m - self.degrees[i]
                monomial = self.get_monomials_of_certain_degree(degree)
                row_coefficients.append(monomial)
            else:
                divisible.append(self.variables[i - 1] **
                                 self.degrees[i - 1])
                degree = self.degree_m - self.degrees[i]
                poss_rows = self.get_monomials_of_certain_degree(degree)
                for div in divisible:
                    for p in poss_rows:
                        if rem(p, div) == 0:
                            poss_rows = [item for item in poss_rows
                                         if item != p]
                row_coefficients.append(poss_rows)
        return row_coefficients

    def get_matrix(self):
        """
        Returns
        =======

        macaulay_matrix: Matrix
            The Macaulay numerator matrix
        """
        rows = []
        row_coefficients = self.get_row_coefficients()
        for i in range(self.n):
            for multiplier in row_coefficients[i]:
                coefficients = []
                poly = Poly(self.polynomials[i] * multiplier,
                            *self.variables)

                for mono in self.monomial_set:
                    coefficients.append(poly.coeff_monomial(mono))
                rows.append(coefficients)

        macaulay_matrix = Matrix(rows)
        return macaulay_matrix

    def get_reduced_nonreduced(self):
        r"""
        Returns
        =======

        reduced: list
            A list of the reduced monomials
        non_reduced: list
            A list of the monomials that are not reduced

        Definition
        ==========

        A polynomial is said to be reduced in x_i, if its degree (the
        maximum degree of its monomials) in x_i is less than d_i. A
        polynomial that is reduced in all variables but one is said
        simply to be reduced.
        """
        divisible = []
        for m in self.monomial_set:
            temp = []
            for i, v in enumerate(self.variables):
                temp.append(bool(total_degree(m, v) >= self.degrees[i]))
            divisible.append(temp)
        reduced = [i for i, r in enumerate(divisible)
                   if sum(r) < self.n - 1]
        non_reduced = [i for i, r in enumerate(divisible)
                       if sum(r) >= self.n -1]

        return reduced, non_reduced

    def get_submatrix(self, matrix):
        r"""
        Returns
        =======

        macaulay_submatrix: Matrix
            The Macaulay denominator matrix. Columns that are non reduced are kept.
            The row which contains one of the a_{i}s is dropped. a_{i}s
            are the coefficients of x_i ^ {d_i}.
        """
        reduced, non_reduced = self.get_reduced_nonreduced()

        # if reduced == [], then det(matrix) should be 1
        if reduced == []:
            return diag([1])

        # reduced != []
        reduction_set = [v ** self.degrees[i] for i, v
                         in enumerate(self.variables)]

        ais = [self.polynomials[i].coeff(reduction_set[i])
               for i in range(self.n)]

        reduced_matrix = matrix[:, reduced]
        keep = []
        for row in range(reduced_matrix.rows):
            check = [ai in reduced_matrix[row, :] for ai in ais]
            if True not in check:
                keep.append(row)

        return matrix[keep, non_reduced]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\worksheet\_reader.py ===
# Copyright (c) 2010-2024 openpyxl

"""Reader for a single worksheet."""
from copy import copy
from warnings import warn

# compatibility imports
from openpyxl.xml.functions import iterparse

# package imports
from openpyxl.cell import Cell, MergedCell
from openpyxl.cell.text import Text
from openpyxl.worksheet.dimensions import (
    ColumnDimension,
    RowDimension,
    SheetFormatProperties,
)

from openpyxl.xml.constants import (
    SHEET_MAIN_NS,
    EXT_TYPES,
)
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.formula.translate import Translator
from openpyxl.utils import (
    get_column_letter,
    coordinate_to_tuple,
    )
from openpyxl.utils.datetime import from_excel, from_ISO8601, WINDOWS_EPOCH
from openpyxl.descriptors.excel import ExtensionList
from openpyxl.cell.rich_text import CellRichText

from .formula import DataTableFormula, ArrayFormula
from .filters import AutoFilter
from .header_footer import HeaderFooter
from .hyperlink import HyperlinkList
from .merge import MergeCells
from .page import PageMargins, PrintOptions, PrintPageSetup
from .pagebreak import RowBreak, ColBreak
from .protection import SheetProtection
from .scenario import ScenarioList
from .views import SheetViewList
from .datavalidation import DataValidationList
from .table import TablePartList
from .properties import WorksheetProperties
from .dimensions import SheetDimension
from .related import Related


CELL_TAG = '{%s}c' % SHEET_MAIN_NS
VALUE_TAG = '{%s}v' % SHEET_MAIN_NS
FORMULA_TAG = '{%s}f' % SHEET_MAIN_NS
MERGE_TAG = '{%s}mergeCells' % SHEET_MAIN_NS
INLINE_STRING = "{%s}is" % SHEET_MAIN_NS
COL_TAG = '{%s}col' % SHEET_MAIN_NS
ROW_TAG = '{%s}row' % SHEET_MAIN_NS
CF_TAG = '{%s}conditionalFormatting' % SHEET_MAIN_NS
LEGACY_TAG = '{%s}legacyDrawing' % SHEET_MAIN_NS
PROT_TAG = '{%s}sheetProtection' % SHEET_MAIN_NS
EXT_TAG = "{%s}extLst" % SHEET_MAIN_NS
HYPERLINK_TAG = "{%s}hyperlinks" % SHEET_MAIN_NS
TABLE_TAG = "{%s}tableParts" % SHEET_MAIN_NS
PRINT_TAG = '{%s}printOptions' % SHEET_MAIN_NS
MARGINS_TAG = '{%s}pageMargins' % SHEET_MAIN_NS
PAGE_TAG = '{%s}pageSetup' % SHEET_MAIN_NS
HEADER_TAG = '{%s}headerFooter' % SHEET_MAIN_NS
FILTER_TAG = '{%s}autoFilter' % SHEET_MAIN_NS
VALIDATION_TAG = '{%s}dataValidations' % SHEET_MAIN_NS
PROPERTIES_TAG = '{%s}sheetPr' % SHEET_MAIN_NS
VIEWS_TAG = '{%s}sheetViews' % SHEET_MAIN_NS
FORMAT_TAG = '{%s}sheetFormatPr' % SHEET_MAIN_NS
ROW_BREAK_TAG = '{%s}rowBreaks' % SHEET_MAIN_NS
COL_BREAK_TAG = '{%s}colBreaks' % SHEET_MAIN_NS
SCENARIOS_TAG = '{%s}scenarios' % SHEET_MAIN_NS
DATA_TAG = '{%s}sheetData' % SHEET_MAIN_NS
DIMENSION_TAG = '{%s}dimension' % SHEET_MAIN_NS
CUSTOM_VIEWS_TAG = '{%s}customSheetViews' % SHEET_MAIN_NS


def _cast_number(value):
    "Convert numbers as string to an int or float"
    if "." in value or "E" in value or "e" in value:
        return float(value)
    return int(value)


def parse_richtext_string(element):
    """
    Parse inline string and preserve rich text formatting
    """
    value = CellRichText.from_tree(element) or ""
    if len(value) == 1 and isinstance(value[0], str):
        value = value[0]
    return value


class WorkSheetParser:

    def __init__(self, src, shared_strings, data_only=False,
                 epoch=WINDOWS_EPOCH, date_formats=set(),
                 timedelta_formats=set(), rich_text=False):
        self.min_row = self.min_col = None
        self.epoch = epoch
        self.source = src
        self.shared_strings = shared_strings
        self.data_only = data_only
        self.shared_formulae = {}
        self.row_counter = self.col_counter = 0
        self.tables = TablePartList()
        self.date_formats = date_formats
        self.timedelta_formats = timedelta_formats
        self.row_dimensions = {}
        self.column_dimensions = {}
        self.number_formats = []
        self.keep_vba = False
        self.hyperlinks = HyperlinkList()
        self.formatting = []
        self.legacy_drawing = None
        self.merged_cells = None
        self.row_breaks = RowBreak()
        self.col_breaks = ColBreak()
        self.rich_text = rich_text


    def parse(self):
        dispatcher = {
            COL_TAG: self.parse_column_dimensions,
            PROT_TAG: self.parse_sheet_protection,
            EXT_TAG: self.parse_extensions,
            CF_TAG: self.parse_formatting,
            LEGACY_TAG: self.parse_legacy,
            ROW_BREAK_TAG: self.parse_row_breaks,
            COL_BREAK_TAG: self.parse_col_breaks,
            CUSTOM_VIEWS_TAG: self.parse_custom_views,
                      }

        properties = {
            PRINT_TAG: ('print_options', PrintOptions),
            MARGINS_TAG: ('page_margins', PageMargins),
            PAGE_TAG: ('page_setup', PrintPageSetup),
            HEADER_TAG: ('HeaderFooter', HeaderFooter),
            FILTER_TAG: ('auto_filter', AutoFilter),
            VALIDATION_TAG: ('data_validations', DataValidationList),
            PROPERTIES_TAG: ('sheet_properties', WorksheetProperties),
            VIEWS_TAG: ('views', SheetViewList),
            FORMAT_TAG: ('sheet_format', SheetFormatProperties),
            SCENARIOS_TAG: ('scenarios', ScenarioList),
            TABLE_TAG: ('tables', TablePartList),
            HYPERLINK_TAG: ('hyperlinks', HyperlinkList),
            MERGE_TAG: ('merged_cells', MergeCells),

        }

        it = iterparse(self.source) # add a finaliser to close the source when this becomes possible

        for _, element in it:
            tag_name = element.tag
            if tag_name in dispatcher:
                dispatcher[tag_name](element)
                element.clear()
            elif tag_name in properties:
                prop = properties[tag_name]
                obj = prop[1].from_tree(element)
                setattr(self, prop[0], obj)
                element.clear()
            elif tag_name == ROW_TAG:
                row = self.parse_row(element)
                element.clear()
                yield row


    def parse_dimensions(self):
        """
        Get worksheet dimensions if they are provided.
        """
        it = iterparse(self.source)

        for _event, element in it:
            if element.tag == DIMENSION_TAG:
                dim = SheetDimension.from_tree(element)
                return dim.boundaries

            elif element.tag == DATA_TAG:
                # Dimensions missing
                break
            element.clear()


    def parse_cell(self, element):
        data_type = element.get('t', 'n')
        coordinate = element.get('r')
        style_id = element.get('s', 0)
        if style_id:
            style_id = int(style_id)

        if data_type == "inlineStr":
            value = None
        else:
            value = element.findtext(VALUE_TAG, None) or None

        if coordinate:
            row, column = coordinate_to_tuple(coordinate)
            self.col_counter = column
        else:
            self.col_counter += 1
            row, column = self.row_counter, self.col_counter

        if not self.data_only and element.find(FORMULA_TAG) is not None:
            data_type = 'f'
            value = self.parse_formula(element)

        elif value is not None:
            if data_type == 'n':
                value = _cast_number(value)
                if style_id in self.date_formats:
                    data_type = 'd'
                    try:
                        value = from_excel(
                            value, self.epoch, timedelta=style_id in self.timedelta_formats
                        )
                    except (OverflowError, ValueError):
                        msg = f"""Cell {coordinate} is marked as a date but the serial value {value} is outside the limits for dates. The cell will be treated as an error."""
                        warn(msg)
                        data_type = "e"
                        value = "#VALUE!"
            elif data_type == 's':
                value = self.shared_strings[int(value)]
            elif data_type == 'b':
                value = bool(int(value))
            elif data_type == "str":
                data_type = "s"
            elif data_type == 'd':
                value = from_ISO8601(value)

        elif data_type == 'inlineStr':
                child = element.find(INLINE_STRING)
                if child is not None:
                    data_type = 's'
                    if self.rich_text:
                        value = parse_richtext_string(child)
                    else:
                        value = Text.from_tree(child).content

        return {'row':row, 'column':column, 'value':value, 'data_type':data_type, 'style_id':style_id}


    def parse_formula(self, element):
        """
        possible formulae types: shared, array, datatable
        """
        formula = element.find(FORMULA_TAG)
        formula_type = formula.get('t')
        coordinate = element.get('r')
        value = "="
        if formula.text is not None:
            value += formula.text

        if formula_type == "array":
            value = ArrayFormula(ref=formula.get('ref'), text=value)

        elif formula_type == "shared":
            idx = formula.get('si')
            if idx in self.shared_formulae:
                trans = self.shared_formulae[idx]
                value = trans.translate_formula(coordinate)
            elif value != "=":
                self.shared_formulae[idx] = Translator(value, coordinate)

        elif formula_type == "dataTable":
            value = DataTableFormula(**formula.attrib)

        return value


    def parse_column_dimensions(self, col):
        attrs = dict(col.attrib)
        column = get_column_letter(int(attrs['min']))
        attrs['index'] = column
        self.column_dimensions[column] = attrs


    def parse_row(self, row):
        attrs = dict(row.attrib)

        if "r" in attrs:
            try:
                self.row_counter = int(attrs['r'])
            except ValueError:
                val = float(attrs['r'])
                if val.is_integer():
                    self.row_counter = int(val)
                else:
                    raise ValueError(f"{attrs['r']} is not a valid row number")
        else:
            self.row_counter += 1
        self.col_counter = 0

        keys = {k for k in attrs if not k.startswith('{')}
        if keys - {'r', 'spans'}:
            # don't create dimension objects unless they have relevant information
            self.row_dimensions[str(self.row_counter)] = attrs

        cells = [self.parse_cell(el) for el in row]
        return self.row_counter, cells


    def parse_formatting(self, element):
        try:
            cf = ConditionalFormatting.from_tree(element)
            self.formatting.append(cf)
        except TypeError as e:
            msg = f"Failed to load a conditional formatting rule. It will be discarded. Cause: {e}"
            warn(msg)


    def parse_sheet_protection(self, element):
        protection = SheetProtection.from_tree(element)
        password = element.get("password")
        if password is not None:
            protection.set_password(password, True)
        self.protection = protection


    def parse_extensions(self, element):
        extLst = ExtensionList.from_tree(element)
        for e in extLst.ext:
            ext_type = EXT_TYPES.get(e.uri.upper(), "Unknown")
            msg = "{0} extension is not supported and will be removed".format(ext_type)
            warn(msg)


    def parse_legacy(self, element):
        obj = Related.from_tree(element)
        self.legacy_drawing = obj.id


    def parse_row_breaks(self, element):
        brk = RowBreak.from_tree(element)
        self.row_breaks = brk


    def parse_col_breaks(self, element):
        brk = ColBreak.from_tree(element)
        self.col_breaks = brk


    def parse_custom_views(self, element):
        # clear page_breaks to avoid duplication which Excel doesn't like
        # basically they're ignored in custom views
        self.row_breaks = RowBreak()
        self.col_breaks = ColBreak()


class WorksheetReader:
    """
    Create a parser and apply it to a workbook
    """

    def __init__(self, ws, xml_source, shared_strings, data_only, rich_text):
        self.ws = ws
        self.parser = WorkSheetParser(xml_source, shared_strings,
                data_only, ws.parent.epoch, ws.parent._date_formats,
                ws.parent._timedelta_formats, rich_text)
        self.tables = []


    def bind_cells(self):
        for idx, row in self.parser.parse():
            for cell in row:
                style = self.ws.parent._cell_styles[cell['style_id']]
                c = Cell(self.ws, row=cell['row'], column=cell['column'], style_array=style)
                c._value = cell['value']
                c.data_type = cell['data_type']
                self.ws._cells[(cell['row'], cell['column'])] = c

        if self.ws._cells:
            self.ws._current_row = self.ws.max_row # use cells not row dimensions


    def bind_formatting(self):
        for cf in self.parser.formatting:
            for rule in cf.rules:
                if rule.dxfId is not None:
                    rule.dxf = self.ws.parent._differential_styles[rule.dxfId]
                self.ws.conditional_formatting[cf] = rule


    def bind_tables(self):
        for t in self.parser.tables.tablePart:
            rel = self.ws._rels.get(t.id)
            self.tables.append(rel.Target)


    def bind_merged_cells(self):
        from openpyxl.worksheet.cell_range import MultiCellRange
        from openpyxl.worksheet.merge import MergedCellRange
        if not self.parser.merged_cells:
            return

        ranges = []
        for cr in self.parser.merged_cells.mergeCell:
            mcr = MergedCellRange(self.ws, cr.ref)
            self.ws._clean_merge_range(mcr)
            ranges.append(mcr)
        self.ws.merged_cells = MultiCellRange(ranges)


    def bind_hyperlinks(self):
        for link in self.parser.hyperlinks.hyperlink:
            if link.id:
                rel = self.ws._rels.get(link.id)
                link.target = rel.Target
            if ":" in link.ref:
                # range of cells
                for row in self.ws[link.ref]:
                    for cell in row:
                        try:
                            cell.hyperlink = copy(link)
                        except AttributeError:
                            pass
            else:
                cell = self.ws[link.ref]
                if isinstance(cell, MergedCell):
                    cell = self.normalize_merged_cell_link(cell.coordinate)
                cell.hyperlink = link

    def normalize_merged_cell_link(self, coord):
        """
        Returns the appropriate cell to which a hyperlink, which references a merged cell at the specified coordinates,
        should be bound.
        """
        for rng in self.ws.merged_cells:
            if coord in rng:
                return self.ws.cell(*rng.top[0])

    def bind_col_dimensions(self):
        for col, cd in self.parser.column_dimensions.items():
            if 'style' in cd:
                key = int(cd['style'])
                cd['style'] = self.ws.parent._cell_styles[key]
            self.ws.column_dimensions[col] = ColumnDimension(self.ws, **cd)


    def bind_row_dimensions(self):
        for row, rd in self.parser.row_dimensions.items():
            if 's' in rd:
                key = int(rd['s'])
                rd['s'] = self.ws.parent._cell_styles[key]
            self.ws.row_dimensions[int(row)] = RowDimension(self.ws, **rd)


    def bind_properties(self):
        for k in ('print_options', 'page_margins', 'page_setup',
                  'HeaderFooter', 'auto_filter', 'data_validations',
                  'sheet_properties', 'views', 'sheet_format',
                  'row_breaks', 'col_breaks', 'scenarios', 'legacy_drawing',
                  'protection',
                  ):
            v = getattr(self.parser, k, None)
            if v is not None:
                setattr(self.ws, k, v)


    def bind_all(self):
        self.bind_cells()
        self.bind_merged_cells()
        self.bind_hyperlinks()
        self.bind_formatting()
        self.bind_col_dimensions()
        self.bind_row_dimensions()
        self.bind_tables()
        self.bind_properties()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\event\base.py ===
# event/base.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

"""Base implementation classes.

The public-facing ``Events`` serves as the base class for an event interface;
its public attributes represent different kinds of events.   These attributes
are mirrored onto a ``_Dispatch`` class, which serves as a container for
collections of listener functions.   These collections are represented both
at the class level of a particular ``_Dispatch`` class as well as within
instances of ``_Dispatch``.

"""
from __future__ import annotations

import typing
from typing import Any
from typing import cast
from typing import Dict
from typing import Generic
from typing import Iterator
from typing import List
from typing import Mapping
from typing import MutableMapping
from typing import Optional
from typing import overload
from typing import Tuple
from typing import Type
from typing import Union
import weakref

from .attr import _ClsLevelDispatch
from .attr import _EmptyListener
from .attr import _InstanceLevelDispatch
from .attr import _JoinedListener
from .registry import _ET
from .registry import _EventKey
from .. import util
from ..util.typing import Literal

_registrars: MutableMapping[str, List[Type[_HasEventsDispatch[Any]]]] = (
    util.defaultdict(list)
)


def _is_event_name(name: str) -> bool:
    # _sa_event prefix is special to support internal-only event names.
    # most event names are just plain method names that aren't
    # underscored.

    return (
        not name.startswith("_") and name != "dispatch"
    ) or name.startswith("_sa_event")


class _UnpickleDispatch:
    """Serializable callable that re-generates an instance of
    :class:`_Dispatch` given a particular :class:`.Events` subclass.

    """

    def __call__(self, _instance_cls: Type[_ET]) -> _Dispatch[_ET]:
        for cls in _instance_cls.__mro__:
            if "dispatch" in cls.__dict__:
                return cast(
                    "_Dispatch[_ET]", cls.__dict__["dispatch"].dispatch
                )._for_class(_instance_cls)
        else:
            raise AttributeError("No class with a 'dispatch' member present.")


class _DispatchCommon(Generic[_ET]):
    __slots__ = ()

    _instance_cls: Optional[Type[_ET]]

    def _join(self, other: _DispatchCommon[_ET]) -> _JoinedDispatcher[_ET]:
        raise NotImplementedError()

    def __getattr__(self, name: str) -> _InstanceLevelDispatch[_ET]:
        raise NotImplementedError()

    @property
    def _events(self) -> Type[_HasEventsDispatch[_ET]]:
        raise NotImplementedError()


class _Dispatch(_DispatchCommon[_ET]):
    """Mirror the event listening definitions of an Events class with
    listener collections.

    Classes which define a "dispatch" member will return a
    non-instantiated :class:`._Dispatch` subclass when the member
    is accessed at the class level.  When the "dispatch" member is
    accessed at the instance level of its owner, an instance
    of the :class:`._Dispatch` class is returned.

    A :class:`._Dispatch` class is generated for each :class:`.Events`
    class defined, by the :meth:`._HasEventsDispatch._create_dispatcher_class`
    method.  The original :class:`.Events` classes remain untouched.
    This decouples the construction of :class:`.Events` subclasses from
    the implementation used by the event internals, and allows
    inspecting tools like Sphinx to work in an unsurprising
    way against the public API.

    """

    # "active_history" is an ORM case we add here.   ideally a better
    # system would be in place for ad-hoc attributes.
    __slots__ = "_parent", "_instance_cls", "__dict__", "_empty_listeners"

    _active_history: bool

    _empty_listener_reg: MutableMapping[
        Type[_ET], Dict[str, _EmptyListener[_ET]]
    ] = weakref.WeakKeyDictionary()

    _empty_listeners: Dict[str, _EmptyListener[_ET]]

    _event_names: List[str]

    _instance_cls: Optional[Type[_ET]]

    _joined_dispatch_cls: Type[_JoinedDispatcher[_ET]]

    _events: Type[_HasEventsDispatch[_ET]]
    """reference back to the Events class.

    Bidirectional against _HasEventsDispatch.dispatch

    """

    def __init__(
        self,
        parent: Optional[_Dispatch[_ET]],
        instance_cls: Optional[Type[_ET]] = None,
    ):
        self._parent = parent
        self._instance_cls = instance_cls

        if instance_cls:
            assert parent is not None
            try:
                self._empty_listeners = self._empty_listener_reg[instance_cls]
            except KeyError:
                self._empty_listeners = self._empty_listener_reg[
                    instance_cls
                ] = {
                    ls.name: _EmptyListener(ls, instance_cls)
                    for ls in parent._event_descriptors
                }
        else:
            self._empty_listeners = {}

    def __getattr__(self, name: str) -> _InstanceLevelDispatch[_ET]:
        # Assign EmptyListeners as attributes on demand
        # to reduce startup time for new dispatch objects.
        try:
            ls = self._empty_listeners[name]
        except KeyError:
            raise AttributeError(name)
        else:
            setattr(self, ls.name, ls)
            return ls

    @property
    def _event_descriptors(self) -> Iterator[_ClsLevelDispatch[_ET]]:
        for k in self._event_names:
            # Yield _ClsLevelDispatch related
            # to relevant event name.
            yield getattr(self, k)

    def _listen(self, event_key: _EventKey[_ET], **kw: Any) -> None:
        return self._events._listen(event_key, **kw)

    def _for_class(self, instance_cls: Type[_ET]) -> _Dispatch[_ET]:
        return self.__class__(self, instance_cls)

    def _for_instance(self, instance: _ET) -> _Dispatch[_ET]:
        instance_cls = instance.__class__
        return self._for_class(instance_cls)

    def _join(self, other: _DispatchCommon[_ET]) -> _JoinedDispatcher[_ET]:
        """Create a 'join' of this :class:`._Dispatch` and another.

        This new dispatcher will dispatch events to both
        :class:`._Dispatch` objects.

        """
        assert "_joined_dispatch_cls" in self.__class__.__dict__

        return self._joined_dispatch_cls(self, other)

    def __reduce__(self) -> Union[str, Tuple[Any, ...]]:
        return _UnpickleDispatch(), (self._instance_cls,)

    def _update(
        self, other: _Dispatch[_ET], only_propagate: bool = True
    ) -> None:
        """Populate from the listeners in another :class:`_Dispatch`
        object."""
        for ls in other._event_descriptors:
            if isinstance(ls, _EmptyListener):
                continue
            getattr(self, ls.name).for_modify(self)._update(
                ls, only_propagate=only_propagate
            )

    def _clear(self) -> None:
        for ls in self._event_descriptors:
            ls.for_modify(self).clear()


def _remove_dispatcher(cls: Type[_HasEventsDispatch[_ET]]) -> None:
    for k in cls.dispatch._event_names:
        _registrars[k].remove(cls)
        if not _registrars[k]:
            del _registrars[k]


class _HasEventsDispatch(Generic[_ET]):
    _dispatch_target: Optional[Type[_ET]]
    """class which will receive the .dispatch collection"""

    dispatch: _Dispatch[_ET]
    """reference back to the _Dispatch class.

    Bidirectional against _Dispatch._events

    """

    if typing.TYPE_CHECKING:

        def __getattr__(self, name: str) -> _InstanceLevelDispatch[_ET]: ...

    def __init_subclass__(cls) -> None:
        """Intercept new Event subclasses and create associated _Dispatch
        classes."""

        cls._create_dispatcher_class(cls.__name__, cls.__bases__, cls.__dict__)

    @classmethod
    def _accept_with(
        cls, target: Union[_ET, Type[_ET]], identifier: str
    ) -> Optional[Union[_ET, Type[_ET]]]:
        raise NotImplementedError()

    @classmethod
    def _listen(
        cls,
        event_key: _EventKey[_ET],
        *,
        propagate: bool = False,
        insert: bool = False,
        named: bool = False,
        asyncio: bool = False,
    ) -> None:
        raise NotImplementedError()

    @staticmethod
    def _set_dispatch(
        klass: Type[_HasEventsDispatch[_ET]],
        dispatch_cls: Type[_Dispatch[_ET]],
    ) -> _Dispatch[_ET]:
        # This allows an Events subclass to define additional utility
        # methods made available to the target via
        # "self.dispatch._events.<utilitymethod>"
        # @staticmethod to allow easy "super" calls while in a metaclass
        # constructor.
        klass.dispatch = dispatch_cls(None)
        dispatch_cls._events = klass
        return klass.dispatch

    @classmethod
    def _create_dispatcher_class(
        cls, classname: str, bases: Tuple[type, ...], dict_: Mapping[str, Any]
    ) -> None:
        """Create a :class:`._Dispatch` class corresponding to an
        :class:`.Events` class."""

        # there's all kinds of ways to do this,
        # i.e. make a Dispatch class that shares the '_listen' method
        # of the Event class, this is the straight monkeypatch.
        if hasattr(cls, "dispatch"):
            dispatch_base = cls.dispatch.__class__
        else:
            dispatch_base = _Dispatch

        event_names = [k for k in dict_ if _is_event_name(k)]
        dispatch_cls = cast(
            "Type[_Dispatch[_ET]]",
            type(
                "%sDispatch" % classname,
                (dispatch_base,),
                {"__slots__": event_names},
            ),
        )

        dispatch_cls._event_names = event_names
        dispatch_inst = cls._set_dispatch(cls, dispatch_cls)
        for k in dispatch_cls._event_names:
            setattr(dispatch_inst, k, _ClsLevelDispatch(cls, dict_[k]))
            _registrars[k].append(cls)

        for super_ in dispatch_cls.__bases__:
            if issubclass(super_, _Dispatch) and super_ is not _Dispatch:
                for ls in super_._events.dispatch._event_descriptors:
                    setattr(dispatch_inst, ls.name, ls)
                    dispatch_cls._event_names.append(ls.name)

        if getattr(cls, "_dispatch_target", None):
            dispatch_target_cls = cls._dispatch_target
            assert dispatch_target_cls is not None
            if (
                hasattr(dispatch_target_cls, "__slots__")
                and "_slots_dispatch" in dispatch_target_cls.__slots__
            ):
                dispatch_target_cls.dispatch = slots_dispatcher(cls)
            else:
                dispatch_target_cls.dispatch = dispatcher(cls)

        klass = type(
            "Joined%s" % dispatch_cls.__name__,
            (_JoinedDispatcher,),
            {"__slots__": event_names},
        )
        dispatch_cls._joined_dispatch_cls = klass

        # establish pickle capability by adding it to this module
        globals()[klass.__name__] = klass


class _JoinedDispatcher(_DispatchCommon[_ET]):
    """Represent a connection between two _Dispatch objects."""

    __slots__ = "local", "parent", "_instance_cls"

    local: _DispatchCommon[_ET]
    parent: _DispatchCommon[_ET]
    _instance_cls: Optional[Type[_ET]]

    def __init__(
        self, local: _DispatchCommon[_ET], parent: _DispatchCommon[_ET]
    ):
        self.local = local
        self.parent = parent
        self._instance_cls = self.local._instance_cls

    def __reduce__(self) -> Any:
        return (self.__class__, (self.local, self.parent))

    def __getattr__(self, name: str) -> _JoinedListener[_ET]:
        # Assign _JoinedListeners as attributes on demand
        # to reduce startup time for new dispatch objects.
        ls = getattr(self.local, name)
        jl = _JoinedListener(self.parent, ls.name, ls)
        setattr(self, ls.name, jl)
        return jl

    def _listen(self, event_key: _EventKey[_ET], **kw: Any) -> None:
        return self.parent._listen(event_key, **kw)

    @property
    def _events(self) -> Type[_HasEventsDispatch[_ET]]:
        return self.parent._events


class Events(_HasEventsDispatch[_ET]):
    """Define event listening functions for a particular target type."""

    @classmethod
    def _accept_with(
        cls, target: Union[_ET, Type[_ET]], identifier: str
    ) -> Optional[Union[_ET, Type[_ET]]]:
        def dispatch_is(*types: Type[Any]) -> bool:
            return all(isinstance(target.dispatch, t) for t in types)

        def dispatch_parent_is(t: Type[Any]) -> bool:
            parent = cast("_JoinedDispatcher[_ET]", target.dispatch).parent
            while isinstance(parent, _JoinedDispatcher):
                parent = cast("_JoinedDispatcher[_ET]", parent).parent

            return isinstance(parent, t)

        # Mapper, ClassManager, Session override this to
        # also accept classes, scoped_sessions, sessionmakers, etc.
        if hasattr(target, "dispatch"):
            if (
                dispatch_is(cls.dispatch.__class__)
                or dispatch_is(type, cls.dispatch.__class__)
                or (
                    dispatch_is(_JoinedDispatcher)
                    and dispatch_parent_is(cls.dispatch.__class__)
                )
            ):
                return target

        return None

    @classmethod
    def _listen(
        cls,
        event_key: _EventKey[_ET],
        *,
        propagate: bool = False,
        insert: bool = False,
        named: bool = False,
        asyncio: bool = False,
    ) -> None:
        event_key.base_listen(
            propagate=propagate, insert=insert, named=named, asyncio=asyncio
        )

    @classmethod
    def _remove(cls, event_key: _EventKey[_ET]) -> None:
        event_key.remove()

    @classmethod
    def _clear(cls) -> None:
        cls.dispatch._clear()


class dispatcher(Generic[_ET]):
    """Descriptor used by target classes to
    deliver the _Dispatch class at the class level
    and produce new _Dispatch instances for target
    instances.

    """

    def __init__(self, events: Type[_HasEventsDispatch[_ET]]):
        self.dispatch = events.dispatch
        self.events = events

    @overload
    def __get__(
        self, obj: Literal[None], cls: Type[Any]
    ) -> Type[_Dispatch[_ET]]: ...

    @overload
    def __get__(self, obj: Any, cls: Type[Any]) -> _DispatchCommon[_ET]: ...

    def __get__(self, obj: Any, cls: Type[Any]) -> Any:
        if obj is None:
            return self.dispatch

        disp = self.dispatch._for_instance(obj)
        try:
            obj.__dict__["dispatch"] = disp
        except AttributeError as ae:
            raise TypeError(
                "target %r doesn't have __dict__, should it be "
                "defining _slots_dispatch?" % (obj,)
            ) from ae
        return disp


class slots_dispatcher(dispatcher[_ET]):
    def __get__(self, obj: Any, cls: Type[Any]) -> Any:
        if obj is None:
            return self.dispatch

        if hasattr(obj, "_slots_dispatch"):
            return obj._slots_dispatch

        disp = self.dispatch._for_instance(obj)
        obj._slots_dispatch = disp
        return disp

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\werkzeug\_reloader.py ===
from __future__ import annotations

import fnmatch
import os
import subprocess
import sys
import threading
import time
import typing as t
from itertools import chain
from pathlib import PurePath

from ._internal import _log

# The various system prefixes where imports are found. Base values are
# different when running in a virtualenv. All reloaders will ignore the
# base paths (usually the system installation). The stat reloader won't
# scan the virtualenv paths, it will only include modules that are
# already imported.
_ignore_always = tuple({sys.base_prefix, sys.base_exec_prefix})
prefix = {*_ignore_always, sys.prefix, sys.exec_prefix}

if hasattr(sys, "real_prefix"):
    # virtualenv < 20
    prefix.add(sys.real_prefix)

_stat_ignore_scan = tuple(prefix)
del prefix
_ignore_common_dirs = {
    "__pycache__",
    ".git",
    ".hg",
    ".tox",
    ".nox",
    ".pytest_cache",
    ".mypy_cache",
}


def _iter_module_paths() -> t.Iterator[str]:
    """Find the filesystem paths associated with imported modules."""
    # List is in case the value is modified by the app while updating.
    for module in list(sys.modules.values()):
        name = getattr(module, "__file__", None)

        if name is None or name.startswith(_ignore_always):
            continue

        while not os.path.isfile(name):
            # Zip file, find the base file without the module path.
            old = name
            name = os.path.dirname(name)

            if name == old:  # skip if it was all directories somehow
                break
        else:
            yield name


def _remove_by_pattern(paths: set[str], exclude_patterns: set[str]) -> None:
    for pattern in exclude_patterns:
        paths.difference_update(fnmatch.filter(paths, pattern))


def _find_stat_paths(
    extra_files: set[str], exclude_patterns: set[str]
) -> t.Iterable[str]:
    """Find paths for the stat reloader to watch. Returns imported
    module files, Python files under non-system paths. Extra files and
    Python files under extra directories can also be scanned.

    System paths have to be excluded for efficiency. Non-system paths,
    such as a project root or ``sys.path.insert``, should be the paths
    of interest to the user anyway.
    """
    paths = set()

    for path in chain(list(sys.path), extra_files):
        path = os.path.abspath(path)

        if os.path.isfile(path):
            # zip file on sys.path, or extra file
            paths.add(path)
            continue

        parent_has_py = {os.path.dirname(path): True}

        for root, dirs, files in os.walk(path):
            # Optimizations: ignore system prefixes, __pycache__ will
            # have a py or pyc module at the import path, ignore some
            # common known dirs such as version control and tool caches.
            if (
                root.startswith(_stat_ignore_scan)
                or os.path.basename(root) in _ignore_common_dirs
            ):
                dirs.clear()
                continue

            has_py = False

            for name in files:
                if name.endswith((".py", ".pyc")):
                    has_py = True
                    paths.add(os.path.join(root, name))

            # Optimization: stop scanning a directory if neither it nor
            # its parent contained Python files.
            if not (has_py or parent_has_py[os.path.dirname(root)]):
                dirs.clear()
                continue

            parent_has_py[root] = has_py

    paths.update(_iter_module_paths())
    _remove_by_pattern(paths, exclude_patterns)
    return paths


def _find_watchdog_paths(
    extra_files: set[str], exclude_patterns: set[str]
) -> t.Iterable[str]:
    """Find paths for the stat reloader to watch. Looks at the same
    sources as the stat reloader, but watches everything under
    directories instead of individual files.
    """
    dirs = set()

    for name in chain(list(sys.path), extra_files):
        name = os.path.abspath(name)

        if os.path.isfile(name):
            name = os.path.dirname(name)

        dirs.add(name)

    for name in _iter_module_paths():
        dirs.add(os.path.dirname(name))

    _remove_by_pattern(dirs, exclude_patterns)
    return _find_common_roots(dirs)


def _find_common_roots(paths: t.Iterable[str]) -> t.Iterable[str]:
    root: dict[str, dict[str, t.Any]] = {}

    for chunks in sorted((PurePath(x).parts for x in paths), key=len, reverse=True):
        node = root

        for chunk in chunks:
            node = node.setdefault(chunk, {})

        node.clear()

    rv = set()

    def _walk(node: t.Mapping[str, dict[str, t.Any]], path: tuple[str, ...]) -> None:
        for prefix, child in node.items():
            _walk(child, path + (prefix,))

        # If there are no more nodes, and a path has been accumulated, add it.
        # Path may be empty if the "" entry is in sys.path.
        if not node and path:
            rv.add(os.path.join(*path))

    _walk(root, ())
    return rv


def _get_args_for_reloading() -> list[str]:
    """Determine how the script was executed, and return the args needed
    to execute it again in a new process.
    """
    if sys.version_info >= (3, 10):
        # sys.orig_argv, added in Python 3.10, contains the exact args used to invoke
        # Python. Still replace argv[0] with sys.executable for accuracy.
        return [sys.executable, *sys.orig_argv[1:]]

    rv = [sys.executable]
    py_script = sys.argv[0]
    args = sys.argv[1:]
    # Need to look at main module to determine how it was executed.
    __main__ = sys.modules["__main__"]

    # The value of __package__ indicates how Python was called. It may
    # not exist if a setuptools script is installed as an egg. It may be
    # set incorrectly for entry points created with pip on Windows.
    if getattr(__main__, "__package__", None) is None or (
        os.name == "nt"
        and __main__.__package__ == ""
        and not os.path.exists(py_script)
        and os.path.exists(f"{py_script}.exe")
    ):
        # Executed a file, like "python app.py".
        py_script = os.path.abspath(py_script)

        if os.name == "nt":
            # Windows entry points have ".exe" extension and should be
            # called directly.
            if not os.path.exists(py_script) and os.path.exists(f"{py_script}.exe"):
                py_script += ".exe"

            if (
                os.path.splitext(sys.executable)[1] == ".exe"
                and os.path.splitext(py_script)[1] == ".exe"
            ):
                rv.pop(0)

        rv.append(py_script)
    else:
        # Executed a module, like "python -m werkzeug.serving".
        if os.path.isfile(py_script):
            # Rewritten by Python from "-m script" to "/path/to/script.py".
            py_module = t.cast(str, __main__.__package__)
            name = os.path.splitext(os.path.basename(py_script))[0]

            if name != "__main__":
                py_module += f".{name}"
        else:
            # Incorrectly rewritten by pydevd debugger from "-m script" to "script".
            py_module = py_script

        rv.extend(("-m", py_module.lstrip(".")))

    rv.extend(args)
    return rv


class ReloaderLoop:
    name = ""

    def __init__(
        self,
        extra_files: t.Iterable[str] | None = None,
        exclude_patterns: t.Iterable[str] | None = None,
        interval: int | float = 1,
    ) -> None:
        self.extra_files: set[str] = {os.path.abspath(x) for x in extra_files or ()}
        self.exclude_patterns: set[str] = set(exclude_patterns or ())
        self.interval = interval

    def __enter__(self) -> ReloaderLoop:
        """Do any setup, then run one step of the watch to populate the
        initial filesystem state.
        """
        self.run_step()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):  # type: ignore
        """Clean up any resources associated with the reloader."""
        pass

    def run(self) -> None:
        """Continually run the watch step, sleeping for the configured
        interval after each step.
        """
        while True:
            self.run_step()
            time.sleep(self.interval)

    def run_step(self) -> None:
        """Run one step for watching the filesystem. Called once to set
        up initial state, then repeatedly to update it.
        """
        pass

    def restart_with_reloader(self) -> int:
        """Spawn a new Python interpreter with the same arguments as the
        current one, but running the reloader thread.
        """
        while True:
            _log("info", f" * Restarting with {self.name}")
            args = _get_args_for_reloading()
            new_environ = os.environ.copy()
            new_environ["WERKZEUG_RUN_MAIN"] = "true"
            exit_code = subprocess.call(args, env=new_environ, close_fds=False)

            if exit_code != 3:
                return exit_code

    def trigger_reload(self, filename: str) -> None:
        self.log_reload(filename)
        sys.exit(3)

    def log_reload(self, filename: str | bytes) -> None:
        filename = os.path.abspath(filename)
        _log("info", f" * Detected change in {filename!r}, reloading")


class StatReloaderLoop(ReloaderLoop):
    name = "stat"

    def __enter__(self) -> ReloaderLoop:
        self.mtimes: dict[str, float] = {}
        return super().__enter__()

    def run_step(self) -> None:
        for name in _find_stat_paths(self.extra_files, self.exclude_patterns):
            try:
                mtime = os.stat(name).st_mtime
            except OSError:
                continue

            old_time = self.mtimes.get(name)

            if old_time is None:
                self.mtimes[name] = mtime
                continue

            if mtime > old_time:
                self.trigger_reload(name)


class WatchdogReloaderLoop(ReloaderLoop):
    def __init__(self, *args: t.Any, **kwargs: t.Any) -> None:
        from watchdog.events import EVENT_TYPE_CLOSED
        from watchdog.events import EVENT_TYPE_CREATED
        from watchdog.events import EVENT_TYPE_DELETED
        from watchdog.events import EVENT_TYPE_MODIFIED
        from watchdog.events import EVENT_TYPE_MOVED
        from watchdog.events import FileModifiedEvent
        from watchdog.events import PatternMatchingEventHandler
        from watchdog.observers import Observer

        super().__init__(*args, **kwargs)
        trigger_reload = self.trigger_reload

        class EventHandler(PatternMatchingEventHandler):
            def on_any_event(self, event: FileModifiedEvent):  # type: ignore
                if event.event_type not in {
                    EVENT_TYPE_CLOSED,
                    EVENT_TYPE_CREATED,
                    EVENT_TYPE_DELETED,
                    EVENT_TYPE_MODIFIED,
                    EVENT_TYPE_MOVED,
                }:
                    # skip events that don't involve changes to the file
                    return

                trigger_reload(event.src_path)

        reloader_name = Observer.__name__.lower()  # type: ignore[attr-defined]

        if reloader_name.endswith("observer"):
            reloader_name = reloader_name[:-8]

        self.name = f"watchdog ({reloader_name})"
        self.observer = Observer()
        # Extra patterns can be non-Python files, match them in addition
        # to all Python files in default and extra directories. Ignore
        # __pycache__ since a change there will always have a change to
        # the source file (or initial pyc file) as well. Ignore Git and
        # Mercurial internal changes.
        extra_patterns = [p for p in self.extra_files if not os.path.isdir(p)]
        self.event_handler = EventHandler(
            patterns=["*.py", "*.pyc", "*.zip", *extra_patterns],
            ignore_patterns=[
                *[f"*/{d}/*" for d in _ignore_common_dirs],
                *self.exclude_patterns,
            ],
        )
        self.should_reload = False

    def trigger_reload(self, filename: str | bytes) -> None:
        # This is called inside an event handler, which means throwing
        # SystemExit has no effect.
        # https://github.com/gorakhargosh/watchdog/issues/294
        self.should_reload = True
        self.log_reload(filename)

    def __enter__(self) -> ReloaderLoop:
        self.watches: dict[str, t.Any] = {}
        self.observer.start()
        return super().__enter__()

    def __exit__(self, exc_type, exc_val, exc_tb):  # type: ignore
        self.observer.stop()
        self.observer.join()

    def run(self) -> None:
        while not self.should_reload:
            self.run_step()
            time.sleep(self.interval)

        sys.exit(3)

    def run_step(self) -> None:
        to_delete = set(self.watches)

        for path in _find_watchdog_paths(self.extra_files, self.exclude_patterns):
            if path not in self.watches:
                try:
                    self.watches[path] = self.observer.schedule(
                        self.event_handler, path, recursive=True
                    )
                except OSError:
                    # Clear this path from list of watches We don't want
                    # the same error message showing again in the next
                    # iteration.
                    self.watches[path] = None

            to_delete.discard(path)

        for path in to_delete:
            watch = self.watches.pop(path, None)

            if watch is not None:
                self.observer.unschedule(watch)


reloader_loops: dict[str, type[ReloaderLoop]] = {
    "stat": StatReloaderLoop,
    "watchdog": WatchdogReloaderLoop,
}

try:
    __import__("watchdog.observers")
except ImportError:
    reloader_loops["auto"] = reloader_loops["stat"]
else:
    reloader_loops["auto"] = reloader_loops["watchdog"]


def ensure_echo_on() -> None:
    """Ensure that echo mode is enabled. Some tools such as PDB disable
    it which causes usability issues after a reload."""
    # tcgetattr will fail if stdin isn't a tty
    if sys.stdin is None or not sys.stdin.isatty():
        return

    try:
        import termios
    except ImportError:
        return

    attributes = termios.tcgetattr(sys.stdin)

    if not attributes[3] & termios.ECHO:
        attributes[3] |= termios.ECHO
        termios.tcsetattr(sys.stdin, termios.TCSANOW, attributes)


def run_with_reloader(
    main_func: t.Callable[[], None],
    extra_files: t.Iterable[str] | None = None,
    exclude_patterns: t.Iterable[str] | None = None,
    interval: int | float = 1,
    reloader_type: str = "auto",
) -> None:
    """Run the given function in an independent Python interpreter."""
    import signal

    signal.signal(signal.SIGTERM, lambda *args: sys.exit(0))
    reloader = reloader_loops[reloader_type](
        extra_files=extra_files, exclude_patterns=exclude_patterns, interval=interval
    )

    try:
        if os.environ.get("WERKZEUG_RUN_MAIN") == "true":
            ensure_echo_on()
            t = threading.Thread(target=main_func, args=())
            t.daemon = True

            # Enter the reloader to set up initial state, then start
            # the app thread and reloader update loop.
            with reloader:
                t.start()
                reloader.run()
        else:
            sys.exit(reloader.restart_with_reloader())
    except KeyboardInterrupt:
        pass

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pyreadline3\console\ironpython_console.py ===
# -*- coding: utf-8 -*-
# *****************************************************************************
#       Copyright (C) 2003-2006 Gary Bishop.
#       Copyright (C) 2006-2020 Jorgen Stenarson. <jorgen.stenarson@bostream.nu>
#       Copyright (C) 2020 Bassem Girgis. <brgirgis@gmail.com>
#
#  Distributed under the terms of the BSD License.  The full license is in
#  the file COPYING, distributed as part of this software.
# *****************************************************************************


import os
import re

import IronPythonConsole
import System
from pyreadline3.console.ansi import AnsiState
from pyreadline3.keysyms import (
    make_keyinfo,
    make_KeyPress,
    make_KeyPress_from_keydescr,
    make_keysym,
)
from pyreadline3.logger import log

from .event import Event

"""Cursor control and color for the .NET console.
"""

#
# Ironpython requires a patch to work do:
#
# In file PythonCommandLine.cs patch line:
#    class PythonCommandLine
#    {

# to:
#    public class PythonCommandLine
#    {
#
#
#
# primitive debug printing that won't interfere with the screen

import sys

import clr

clr.AddReferenceToFileAndPath(sys.executable)


color = System.ConsoleColor

ansicolor = {
    "0;30": color.Black,
    "0;31": color.DarkRed,
    "0;32": color.DarkGreen,
    "0;33": color.DarkYellow,
    "0;34": color.DarkBlue,
    "0;35": color.DarkMagenta,
    "0;36": color.DarkCyan,
    "0;37": color.DarkGray,
    "1;30": color.Gray,
    "1;31": color.Red,
    "1;32": color.Green,
    "1;33": color.Yellow,
    "1;34": color.Blue,
    "1;35": color.Magenta,
    "1;36": color.Cyan,
    "1;37": color.White,
}

winattr = {
    "black": 0,
    "darkgray": 0 + 8,
    "darkred": 4,
    "red": 4 + 8,
    "darkgreen": 2,
    "green": 2 + 8,
    "darkyellow": 6,
    "yellow": 6 + 8,
    "darkblue": 1,
    "blue": 1 + 8,
    "darkmagenta": 5,
    "magenta": 5 + 8,
    "darkcyan": 3,
    "cyan": 3 + 8,
    "gray": 7,
    "white": 7 + 8,
}


class Console(object):
    """Console driver for Windows."""

    def __init__(self, newbuffer=0):
        """Initialize the Console object.

        newbuffer=1 will allocate a new buffer so the old content will be restored
        on exit.
        """
        self.serial = 0
        self.attr = System.Console.ForegroundColor
        self.saveattr = winattr[str(System.Console.ForegroundColor).lower()]
        self.savebg = System.Console.BackgroundColor
        log("initial attr=%s" % self.attr)

    def _get(self):
        top = System.Console.WindowTop
        log("WindowTop:%s" % top)
        return top

    def _set(self, value):
        top = System.Console.WindowTop
        log("Set WindowTop:old:%s,new:%s" % (top, value))

    WindowTop = property(_get, _set)
    del _get, _set

    def __del__(self):
        """Cleanup the console when finished."""
        # I don't think this ever gets called
        pass

    def pos(self, x=None, y=None):
        """Move or query the window cursor."""
        if x is not None:
            System.Console.CursorLeft = x
        else:
            x = System.Console.CursorLeft
        if y is not None:
            System.Console.CursorTop = y
        else:
            y = System.Console.CursorTop
        return x, y

    def home(self):
        """Move to home."""
        self.pos(0, 0)

    # Map ANSI color escape sequences into Windows Console Attributes

    terminal_escape = re.compile("(\001?\033\\[[0-9;]*m\002?)")
    escape_parts = re.compile("\001?\033\\[([0-9;]*)m\002?")

    # This pattern should match all characters that change the cursor position differently
    # than a normal character.
    motion_char_re = re.compile("([\n\r\t\010\007])")

    def write_scrolling(self, text, attr=None):
        """write text at current cursor position while watching for scrolling.

        If the window scrolls because you are at the bottom of the screen
        buffer, all positions that you are storing will be shifted by the
        scroll amount. For example, I remember the cursor position of the
        prompt so that I can redraw the line but if the window scrolls,
        the remembered position is off.

        This variant of write tries to keep track of the cursor position
        so that it will know when the screen buffer is scrolled. It
        returns the number of lines that the buffer scrolled.

        """
        x, y = self.pos()
        w, h = self.size()
        scroll = 0  # the result

        # split the string into ordinary characters and funny characters
        chunks = self.motion_char_re.split(text)
        for chunk in chunks:
            n = self.write_color(chunk, attr)
            if len(chunk) == 1:  # the funny characters will be alone
                if chunk[0] == "\n":  # newline
                    x = 0
                    y += 1
                elif chunk[0] == "\r":  # carriage return
                    x = 0
                elif chunk[0] == "\t":  # tab
                    x = 8 * (int(x / 8) + 1)
                    if x > w:  # newline
                        x -= w
                        y += 1
                elif chunk[0] == "\007":  # bell
                    pass
                elif chunk[0] == "\010":
                    x -= 1
                    if x < 0:
                        y -= 1  # backed up 1 line
                else:  # ordinary character
                    x += 1
                if x == w:  # wrap
                    x = 0
                    y += 1
                if y == h:  # scroll
                    scroll += 1
                    y = h - 1
            else:  # chunk of ordinary characters
                x += n
                l = int(x / w)  # lines we advanced
                x = x % w  # new x value
                y += l
                if y >= h:  # scroll
                    scroll += y - h + 1
                    y = h - 1
        return scroll

    trtable = {
        0: color.Black,
        4: color.DarkRed,
        2: color.DarkGreen,
        6: color.DarkYellow,
        1: color.DarkBlue,
        5: color.DarkMagenta,
        3: color.DarkCyan,
        7: color.Gray,
        8: color.DarkGray,
        4 + 8: color.Red,
        2 + 8: color.Green,
        6 + 8: color.Yellow,
        1 + 8: color.Blue,
        5 + 8: color.Magenta,
        3 + 8: color.Cyan,
        7 + 8: color.White,
    }

    def write_color(self, text, attr=None):
        """write text at current cursor position and interpret color escapes.

        return the number of characters written.
        """
        log('write_color("%s", %s)' % (text, attr))
        chunks = self.terminal_escape.split(text)
        log("chunks=%s" % repr(chunks))
        bg = self.savebg
        n = 0  # count the characters we actually write, omitting the escapes
        if attr is None:  # use attribute from initial console
            attr = self.attr
        try:
            fg = self.trtable[(0x000F & attr)]
            bg = self.trtable[(0x00F0 & attr) >> 4]
        except TypeError:
            fg = attr

        for chunk in chunks:
            m = self.escape_parts.match(chunk)
            if m:
                log(m.group(1))
                attr = ansicolor.get(m.group(1), self.attr)
            n += len(chunk)
            System.Console.ForegroundColor = fg
            System.Console.BackgroundColor = bg
            System.Console.Write(chunk)
        return n

    def write_plain(self, text, attr=None):
        """write text at current cursor position."""
        log('write("%s", %s)' % (text, attr))
        if attr is None:
            attr = self.attr
        n = c_int(0)
        self.SetConsoleTextAttribute(self.hout, attr)
        self.WriteConsoleA(self.hout, text, len(text), byref(n), None)
        return len(text)

    if "EMACS" in os.environ:

        def write_color(self, text, attr=None):
            junk = c_int(0)
            self.WriteFile(self.hout, text, len(text), byref(junk), None)
            return len(text)

        write_plain = write_color

    # make this class look like a file object
    def write(self, text):
        log('write("%s")' % text)
        return self.write_color(text)

    # write = write_scrolling

    def isatty(self):
        return True

    def flush(self):
        pass

    def page(self, attr=None, fill=" "):
        """Fill the entire screen."""
        System.Console.Clear()

    def text(self, x, y, text, attr=None):
        """Write text at the given position."""
        self.pos(x, y)
        self.write_color(text, attr)

    def clear_to_end_of_window(self):
        oldtop = self.WindowTop
        lastline = self.WindowTop + System.Console.WindowHeight
        pos = self.pos()
        w, h = self.size()
        length = w - pos[0] + min((lastline - pos[1] - 1), 5) * w - 1
        self.write_color(length * " ")
        self.pos(*pos)
        self.WindowTop = oldtop

    def rectangle(self, rect, attr=None, fill=" "):
        """Fill Rectangle."""
        oldtop = self.WindowTop
        oldpos = self.pos()
        # raise NotImplementedError
        x0, y0, x1, y1 = rect
        if attr is None:
            attr = self.attr
        if fill:
            rowfill = fill[:1] * abs(x1 - x0)
        else:
            rowfill = " " * abs(x1 - x0)
        for y in range(y0, y1):
            System.Console.SetCursorPosition(x0, y)
            self.write_color(rowfill, attr)
        self.pos(*oldpos)

    def scroll(self, rect, dx, dy, attr=None, fill=" "):
        """Scroll a rectangle."""
        raise NotImplementedError

    def scroll_window(self, lines):
        """Scroll the window by the indicated number of lines."""
        top = self.WindowTop + lines
        if top < 0:
            top = 0
        if top + System.Console.WindowHeight > System.Console.BufferHeight:
            top = System.Console.BufferHeight
        self.WindowTop = top

    def getkeypress(self):
        """Return next key press event from the queue, ignoring others."""
        ck = System.ConsoleKey
        while True:
            e = System.Console.ReadKey(True)
            if e.Key == System.ConsoleKey.PageDown:  # PageDown
                self.scroll_window(12)
            elif e.Key == System.ConsoleKey.PageUp:  # PageUp
                self.scroll_window(-12)
            elif str(e.KeyChar) == "\000":  # Drop deadkeys
                log("Deadkey: %s" % e)
                return event(self, e)
            else:
                return event(self, e)

    def title(self, txt=None):
        """Set/get title."""
        if txt:
            System.Console.Title = txt
        else:
            return System.Console.Title

    def size(self, width=None, height=None):
        """Set/get window size."""
        sc = System.Console
        if width is not None and height is not None:
            sc.BufferWidth, sc.BufferHeight = width, height
        else:
            return sc.BufferWidth, sc.BufferHeight

        if width is not None and height is not None:
            sc.WindowWidth, sc.WindowHeight = width, height
        else:
            return sc.WindowWidth - 1, sc.WindowHeight - 1

    def cursor(self, visible=True, size=None):
        """Set cursor on or off."""
        System.Console.CursorVisible = visible

    def bell(self):
        System.Console.Beep()

    def next_serial(self):
        """Get next event serial number."""
        self.serial += 1
        return self.serial


class event(Event):
    """Represent events from the console."""

    def __init__(self, console, input):
        """Initialize an event from the Windows input structure."""
        self.type = "??"
        self.serial = console.next_serial()
        self.width = 0
        self.height = 0
        self.x = 0
        self.y = 0
        self.char = str(input.KeyChar)
        self.keycode = input.Key
        self.state = input.Modifiers
        log("%s,%s,%s" % (input.Modifiers, input.Key, input.KeyChar))
        self.type = "KeyRelease"
        self.keysym = make_keysym(self.keycode)
        self.keyinfo = make_KeyPress(self.char, self.state, self.keycode)


def make_event_from_keydescr(keydescr):
    def input():
        return 1

    input.KeyChar = "a"
    input.Key = System.ConsoleKey.A
    input.Modifiers = System.ConsoleModifiers.Shift
    input.next_serial = input
    e = event(input, input)
    del input.next_serial
    keyinfo = make_KeyPress_from_keydescr(keydescr)
    e.keyinfo = keyinfo
    return e


CTRL_C_EVENT = make_event_from_keydescr("Control-c")


def install_readline(hook):
    def hook_wrap():
        try:
            res = hook()
        except KeyboardInterrupt as x:  # this exception does not seem to be caught
            res = ""
        except EOFError:
            return None
        if res[-1:] == "\n":
            return res[:-1]
        else:
            return res

    class IronPythonWrapper(IronPythonConsole.IConsole):
        def ReadLine(self, autoIndentSize):
            return hook_wrap()

        def Write(self, text, style):
            System.Console.Write(text)

        def WriteLine(self, text, style):
            System.Console.WriteLine(text)

    IronPythonConsole.PythonCommandLine.MyConsole = IronPythonWrapper()


if __name__ == "__main__":
    import sys
    import time

    c = Console(0)
    sys.stdout = c
    sys.stderr = c
    c.page()
    c.pos(5, 10)
    c.write("hi there")
    c.title("Testing console")
    #    c.bell()
    print()
    print("size", c.size())
    print("  some printed output")
    for i in range(10):
        e = c.getkeypress()
        print(e.Key, chr(e.KeyChar), ord(e.KeyChar), e.Modifiers)
    del c

    System.Console.Clear()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\tensor\index_methods.py ===
"""Module with functions operating on IndexedBase, Indexed and Idx objects

    - Check shape conformance
    - Determine indices in resulting expression

    etc.

    Methods in this module could be implemented by calling methods on Expr
    objects instead.  When things stabilize this could be a useful
    refactoring.
"""

from functools import reduce

from sympy.core.function import Function
from sympy.functions import exp, Piecewise
from sympy.tensor.indexed import Idx, Indexed
from sympy.utilities import sift

from collections import OrderedDict

class IndexConformanceException(Exception):
    pass

def _unique_and_repeated(inds):
    """
    Returns the unique and repeated indices. Also note, from the examples given below
    that the order of indices is maintained as given in the input.

    Examples
    ========

    >>> from sympy.tensor.index_methods import _unique_and_repeated
    >>> _unique_and_repeated([2, 3, 1, 3, 0, 4, 0])
    ([2, 1, 4], [3, 0])
    """
    uniq = OrderedDict()
    for i in inds:
        if i in uniq:
            uniq[i] = 0
        else:
            uniq[i] = 1
    return sift(uniq, lambda x: uniq[x], binary=True)

def _remove_repeated(inds):
    """
    Removes repeated objects from sequences

    Returns a set of the unique objects and a tuple of all that have been
    removed.

    Examples
    ========

    >>> from sympy.tensor.index_methods import _remove_repeated
    >>> l1 = [1, 2, 3, 2]
    >>> _remove_repeated(l1)
    ({1, 3}, (2,))

    """
    u, r = _unique_and_repeated(inds)
    return set(u), tuple(r)


def _get_indices_Mul(expr, return_dummies=False):
    """Determine the outer indices of a Mul object.

    Examples
    ========

    >>> from sympy.tensor.index_methods import _get_indices_Mul
    >>> from sympy.tensor.indexed import IndexedBase, Idx
    >>> i, j, k = map(Idx, ['i', 'j', 'k'])
    >>> x = IndexedBase('x')
    >>> y = IndexedBase('y')
    >>> _get_indices_Mul(x[i, k]*y[j, k])
    ({i, j}, {})
    >>> _get_indices_Mul(x[i, k]*y[j, k], return_dummies=True)
    ({i, j}, {}, (k,))

    """

    inds = list(map(get_indices, expr.args))
    inds, syms = list(zip(*inds))

    inds = list(map(list, inds))
    inds = list(reduce(lambda x, y: x + y, inds))
    inds, dummies = _remove_repeated(inds)

    symmetry = {}
    for s in syms:
        for pair in s:
            if pair in symmetry:
                symmetry[pair] *= s[pair]
            else:
                symmetry[pair] = s[pair]

    if return_dummies:
        return inds, symmetry, dummies
    else:
        return inds, symmetry


def _get_indices_Pow(expr):
    """Determine outer indices of a power or an exponential.

    A power is considered a universal function, so that the indices of a Pow is
    just the collection of indices present in the expression.  This may be
    viewed as a bit inconsistent in the special case:

        x[i]**2 = x[i]*x[i]                                                      (1)

    The above expression could have been interpreted as the contraction of x[i]
    with itself, but we choose instead to interpret it as a function

        lambda y: y**2

    applied to each element of x (a universal function in numpy terms).  In
    order to allow an interpretation of (1) as a contraction, we need
    contravariant and covariant Idx subclasses.  (FIXME: this is not yet
    implemented)

    Expressions in the base or exponent are subject to contraction as usual,
    but an index that is present in the exponent, will not be considered
    contractable with its own base.  Note however, that indices in the same
    exponent can be contracted with each other.

    Examples
    ========

    >>> from sympy.tensor.index_methods import _get_indices_Pow
    >>> from sympy import Pow, exp, IndexedBase, Idx
    >>> A = IndexedBase('A')
    >>> x = IndexedBase('x')
    >>> i, j, k = map(Idx, ['i', 'j', 'k'])
    >>> _get_indices_Pow(exp(A[i, j]*x[j]))
    ({i}, {})
    >>> _get_indices_Pow(Pow(x[i], x[i]))
    ({i}, {})
    >>> _get_indices_Pow(Pow(A[i, j]*x[j], x[i]))
    ({i}, {})

    """
    base, exp = expr.as_base_exp()
    binds, bsyms = get_indices(base)
    einds, esyms = get_indices(exp)

    inds = binds | einds

    # FIXME: symmetries from power needs to check special cases, else nothing
    symmetries = {}

    return inds, symmetries


def _get_indices_Add(expr):
    """Determine outer indices of an Add object.

    In a sum, each term must have the same set of outer indices.  A valid
    expression could be

        x(i)*y(j) - x(j)*y(i)

    But we do not allow expressions like:

        x(i)*y(j) - z(j)*z(j)

    FIXME: Add support for Numpy broadcasting

    Examples
    ========

    >>> from sympy.tensor.index_methods import _get_indices_Add
    >>> from sympy.tensor.indexed import IndexedBase, Idx
    >>> i, j, k = map(Idx, ['i', 'j', 'k'])
    >>> x = IndexedBase('x')
    >>> y = IndexedBase('y')
    >>> _get_indices_Add(x[i] + x[k]*y[i, k])
    ({i}, {})

    """

    inds = list(map(get_indices, expr.args))
    inds, syms = list(zip(*inds))

    # allow broadcast of scalars
    non_scalars = [x for x in inds if x != set()]
    if not non_scalars:
        return set(), {}

    if not all(x == non_scalars[0] for x in non_scalars[1:]):
        raise IndexConformanceException("Indices are not consistent: %s" % expr)
    if not reduce(lambda x, y: x != y or y, syms):
        symmetries = syms[0]
    else:
        # FIXME: search for symmetries
        symmetries = {}

    return non_scalars[0], symmetries


def get_indices(expr):
    """Determine the outer indices of expression ``expr``

    By *outer* we mean indices that are not summation indices.  Returns a set
    and a dict.  The set contains outer indices and the dict contains
    information about index symmetries.

    Examples
    ========

    >>> from sympy.tensor.index_methods import get_indices
    >>> from sympy import symbols
    >>> from sympy.tensor import IndexedBase
    >>> x, y, A = map(IndexedBase, ['x', 'y', 'A'])
    >>> i, j, a, z = symbols('i j a z', integer=True)

    The indices of the total expression is determined, Repeated indices imply a
    summation, for instance the trace of a matrix A:

    >>> get_indices(A[i, i])
    (set(), {})

    In the case of many terms, the terms are required to have identical
    outer indices.  Else an IndexConformanceException is raised.

    >>> get_indices(x[i] + A[i, j]*y[j])
    ({i}, {})

    :Exceptions:

    An IndexConformanceException means that the terms ar not compatible, e.g.

    >>> get_indices(x[i] + y[j])                #doctest: +SKIP
            (...)
    IndexConformanceException: Indices are not consistent: x(i) + y(j)

    .. warning::
       The concept of *outer* indices applies recursively, starting on the deepest
       level.  This implies that dummies inside parenthesis are assumed to be
       summed first, so that the following expression is handled gracefully:

       >>> get_indices((x[i] + A[i, j]*y[j])*x[j])
       ({i, j}, {})

       This is correct and may appear convenient, but you need to be careful
       with this as SymPy will happily .expand() the product, if requested.  The
       resulting expression would mix the outer ``j`` with the dummies inside
       the parenthesis, which makes it a different expression.  To be on the
       safe side, it is best to avoid such ambiguities by using unique indices
       for all contractions that should be held separate.

    """
    # We call ourself recursively to determine indices of sub expressions.

    # break recursion
    if isinstance(expr, Indexed):
        c = expr.indices
        inds, dummies = _remove_repeated(c)
        return inds, {}
    elif expr is None:
        return set(), {}
    elif isinstance(expr, Idx):
        return {expr}, {}
    elif expr.is_Atom:
        return set(), {}


    # recurse via specialized functions
    else:
        if expr.is_Mul:
            return _get_indices_Mul(expr)
        elif expr.is_Add:
            return _get_indices_Add(expr)
        elif expr.is_Pow or isinstance(expr, exp):
            return _get_indices_Pow(expr)

        elif isinstance(expr, Piecewise):
            # FIXME:  No support for Piecewise yet
            return set(), {}
        elif isinstance(expr, Function):
            # Support ufunc like behaviour by returning indices from arguments.
            # Functions do not interpret repeated indices across arguments
            # as summation
            ind0 = set()
            for arg in expr.args:
                ind, sym = get_indices(arg)
                ind0 |= ind
            return ind0, sym

        # this test is expensive, so it should be at the end
        elif not expr.has(Indexed):
            return set(), {}
        raise NotImplementedError(
            "FIXME: No specialized handling of type %s" % type(expr))


def get_contraction_structure(expr):
    """Determine dummy indices of ``expr`` and describe its structure

    By *dummy* we mean indices that are summation indices.

    The structure of the expression is determined and described as follows:

    1) A conforming summation of Indexed objects is described with a dict where
       the keys are summation indices and the corresponding values are sets
       containing all terms for which the summation applies.  All Add objects
       in the SymPy expression tree are described like this.

    2) For all nodes in the SymPy expression tree that are *not* of type Add, the
       following applies:

       If a node discovers contractions in one of its arguments, the node
       itself will be stored as a key in the dict.  For that key, the
       corresponding value is a list of dicts, each of which is the result of a
       recursive call to get_contraction_structure().  The list contains only
       dicts for the non-trivial deeper contractions, omitting dicts with None
       as the one and only key.

    .. Note:: The presence of expressions among the dictionary keys indicates
       multiple levels of index contractions.  A nested dict displays nested
       contractions and may itself contain dicts from a deeper level.  In
       practical calculations the summation in the deepest nested level must be
       calculated first so that the outer expression can access the resulting
       indexed object.

    Examples
    ========

    >>> from sympy.tensor.index_methods import get_contraction_structure
    >>> from sympy import default_sort_key
    >>> from sympy.tensor import IndexedBase, Idx
    >>> x, y, A = map(IndexedBase, ['x', 'y', 'A'])
    >>> i, j, k, l = map(Idx, ['i', 'j', 'k', 'l'])
    >>> get_contraction_structure(x[i]*y[i] + A[j, j])
    {(i,): {x[i]*y[i]}, (j,): {A[j, j]}}
    >>> get_contraction_structure(x[i]*y[j])
    {None: {x[i]*y[j]}}

    A multiplication of contracted factors results in nested dicts representing
    the internal contractions.

    >>> d = get_contraction_structure(x[i, i]*y[j, j])
    >>> sorted(d.keys(), key=default_sort_key)
    [None, x[i, i]*y[j, j]]

    In this case, the product has no contractions:

    >>> d[None]
    {x[i, i]*y[j, j]}

    Factors are contracted "first":

    >>> sorted(d[x[i, i]*y[j, j]], key=default_sort_key)
    [{(i,): {x[i, i]}}, {(j,): {y[j, j]}}]

    A parenthesized Add object is also returned as a nested dictionary.  The
    term containing the parenthesis is a Mul with a contraction among the
    arguments, so it will be found as a key in the result.  It stores the
    dictionary resulting from a recursive call on the Add expression.

    >>> d = get_contraction_structure(x[i]*(y[i] + A[i, j]*x[j]))
    >>> sorted(d.keys(), key=default_sort_key)
    [(A[i, j]*x[j] + y[i])*x[i], (i,)]
    >>> d[(i,)]
    {(A[i, j]*x[j] + y[i])*x[i]}
    >>> d[x[i]*(A[i, j]*x[j] + y[i])]
    [{None: {y[i]}, (j,): {A[i, j]*x[j]}}]

    Powers with contractions in either base or exponent will also be found as
    keys in the dictionary, mapping to a list of results from recursive calls:

    >>> d = get_contraction_structure(A[j, j]**A[i, i])
    >>> d[None]
    {A[j, j]**A[i, i]}
    >>> nested_contractions = d[A[j, j]**A[i, i]]
    >>> nested_contractions[0]
    {(j,): {A[j, j]}}
    >>> nested_contractions[1]
    {(i,): {A[i, i]}}

    The description of the contraction structure may appear complicated when
    represented with a string in the above examples, but it is easy to iterate
    over:

    >>> from sympy import Expr
    >>> for key in d:
    ...     if isinstance(key, Expr):
    ...         continue
    ...     for term in d[key]:
    ...         if term in d:
    ...             # treat deepest contraction first
    ...             pass
    ...     # treat outermost contactions here

    """

    # We call ourself recursively to inspect sub expressions.

    if isinstance(expr, Indexed):
        junk, key = _remove_repeated(expr.indices)
        return {key or None: {expr}}
    elif expr.is_Atom:
        return {None: {expr}}
    elif expr.is_Mul:
        junk, junk, key = _get_indices_Mul(expr, return_dummies=True)
        result = {key or None: {expr}}
        # recurse on every factor
        nested = []
        for fac in expr.args:
            facd = get_contraction_structure(fac)
            if not (None in facd and len(facd) == 1):
                nested.append(facd)
        if nested:
            result[expr] = nested
        return result
    elif expr.is_Pow or isinstance(expr, exp):
        # recurse in base and exp separately.  If either has internal
        # contractions we must include ourselves as a key in the returned dict
        b, e = expr.as_base_exp()
        dbase = get_contraction_structure(b)
        dexp = get_contraction_structure(e)

        dicts = []
        for d in dbase, dexp:
            if not (None in d and len(d) == 1):
                dicts.append(d)
        result = {None: {expr}}
        if dicts:
            result[expr] = dicts
        return result
    elif expr.is_Add:
        # Note: we just collect all terms with identical summation indices, We
        # do nothing to identify equivalent terms here, as this would require
        # substitutions or pattern matching in expressions of unknown
        # complexity.
        result = {}
        for term in expr.args:
            # recurse on every term
            d = get_contraction_structure(term)
            for key in d:
                if key in result:
                    result[key] |= d[key]
                else:
                    result[key] = d[key]
        return result

    elif isinstance(expr, Piecewise):
        # FIXME:  No support for Piecewise yet
        return {None: expr}
    elif isinstance(expr, Function):
        # Collect non-trivial contraction structures in each argument
        # We do not report repeated indices in separate arguments as a
        # contraction
        deeplist = []
        for arg in expr.args:
            deep = get_contraction_structure(arg)
            if not (None in deep and len(deep) == 1):
                deeplist.append(deep)
        d = {None: {expr}}
        if deeplist:
            d[expr] = deeplist
        return d

    # this test is expensive, so it should be at the end
    elif not expr.has(Indexed):
        return {None: {expr}}
    raise NotImplementedError(
        "FIXME: No specialized handling of type %s" % type(expr))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\urllib3\util\url.py ===
from __future__ import annotations

import re
import typing

from ..exceptions import LocationParseError
from .util import to_str

# We only want to normalize urls with an HTTP(S) scheme.
# urllib3 infers URLs without a scheme (None) to be http.
_NORMALIZABLE_SCHEMES = ("http", "https", None)

# Almost all of these patterns were derived from the
# 'rfc3986' module: https://github.com/python-hyper/rfc3986
_PERCENT_RE = re.compile(r"%[a-fA-F0-9]{2}")
_SCHEME_RE = re.compile(r"^(?:[a-zA-Z][a-zA-Z0-9+-]*:|/)")
_URI_RE = re.compile(
    r"^(?:([a-zA-Z][a-zA-Z0-9+.-]*):)?"
    r"(?://([^\\/?#]*))?"
    r"([^?#]*)"
    r"(?:\?([^#]*))?"
    r"(?:#(.*))?$",
    re.UNICODE | re.DOTALL,
)

_IPV4_PAT = r"(?:[0-9]{1,3}\.){3}[0-9]{1,3}"
_HEX_PAT = "[0-9A-Fa-f]{1,4}"
_LS32_PAT = "(?:{hex}:{hex}|{ipv4})".format(hex=_HEX_PAT, ipv4=_IPV4_PAT)
_subs = {"hex": _HEX_PAT, "ls32": _LS32_PAT}
_variations = [
    #                            6( h16 ":" ) ls32
    "(?:%(hex)s:){6}%(ls32)s",
    #                       "::" 5( h16 ":" ) ls32
    "::(?:%(hex)s:){5}%(ls32)s",
    # [               h16 ] "::" 4( h16 ":" ) ls32
    "(?:%(hex)s)?::(?:%(hex)s:){4}%(ls32)s",
    # [ *1( h16 ":" ) h16 ] "::" 3( h16 ":" ) ls32
    "(?:(?:%(hex)s:)?%(hex)s)?::(?:%(hex)s:){3}%(ls32)s",
    # [ *2( h16 ":" ) h16 ] "::" 2( h16 ":" ) ls32
    "(?:(?:%(hex)s:){0,2}%(hex)s)?::(?:%(hex)s:){2}%(ls32)s",
    # [ *3( h16 ":" ) h16 ] "::"    h16 ":"   ls32
    "(?:(?:%(hex)s:){0,3}%(hex)s)?::%(hex)s:%(ls32)s",
    # [ *4( h16 ":" ) h16 ] "::"              ls32
    "(?:(?:%(hex)s:){0,4}%(hex)s)?::%(ls32)s",
    # [ *5( h16 ":" ) h16 ] "::"              h16
    "(?:(?:%(hex)s:){0,5}%(hex)s)?::%(hex)s",
    # [ *6( h16 ":" ) h16 ] "::"
    "(?:(?:%(hex)s:){0,6}%(hex)s)?::",
]

_UNRESERVED_PAT = r"ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789._\-~"
_IPV6_PAT = "(?:" + "|".join([x % _subs for x in _variations]) + ")"
_ZONE_ID_PAT = "(?:%25|%)(?:[" + _UNRESERVED_PAT + "]|%[a-fA-F0-9]{2})+"
_IPV6_ADDRZ_PAT = r"\[" + _IPV6_PAT + r"(?:" + _ZONE_ID_PAT + r")?\]"
_REG_NAME_PAT = r"(?:[^\[\]%:/?#]|%[a-fA-F0-9]{2})*"
_TARGET_RE = re.compile(r"^(/[^?#]*)(?:\?([^#]*))?(?:#.*)?$")

_IPV4_RE = re.compile("^" + _IPV4_PAT + "$")
_IPV6_RE = re.compile("^" + _IPV6_PAT + "$")
_IPV6_ADDRZ_RE = re.compile("^" + _IPV6_ADDRZ_PAT + "$")
_BRACELESS_IPV6_ADDRZ_RE = re.compile("^" + _IPV6_ADDRZ_PAT[2:-2] + "$")
_ZONE_ID_RE = re.compile("(" + _ZONE_ID_PAT + r")\]$")

_HOST_PORT_PAT = ("^(%s|%s|%s)(?::0*?(|0|[1-9][0-9]{0,4}))?$") % (
    _REG_NAME_PAT,
    _IPV4_PAT,
    _IPV6_ADDRZ_PAT,
)
_HOST_PORT_RE = re.compile(_HOST_PORT_PAT, re.UNICODE | re.DOTALL)

_UNRESERVED_CHARS = set(
    "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789._-~"
)
_SUB_DELIM_CHARS = set("!$&'()*+,;=")
_USERINFO_CHARS = _UNRESERVED_CHARS | _SUB_DELIM_CHARS | {":"}
_PATH_CHARS = _USERINFO_CHARS | {"@", "/"}
_QUERY_CHARS = _FRAGMENT_CHARS = _PATH_CHARS | {"?"}


class Url(
    typing.NamedTuple(
        "Url",
        [
            ("scheme", typing.Optional[str]),
            ("auth", typing.Optional[str]),
            ("host", typing.Optional[str]),
            ("port", typing.Optional[int]),
            ("path", typing.Optional[str]),
            ("query", typing.Optional[str]),
            ("fragment", typing.Optional[str]),
        ],
    )
):
    """
    Data structure for representing an HTTP URL. Used as a return value for
    :func:`parse_url`. Both the scheme and host are normalized as they are
    both case-insensitive according to RFC 3986.
    """

    def __new__(  # type: ignore[no-untyped-def]
        cls,
        scheme: str | None = None,
        auth: str | None = None,
        host: str | None = None,
        port: int | None = None,
        path: str | None = None,
        query: str | None = None,
        fragment: str | None = None,
    ):
        if path and not path.startswith("/"):
            path = "/" + path
        if scheme is not None:
            scheme = scheme.lower()
        return super().__new__(cls, scheme, auth, host, port, path, query, fragment)

    @property
    def hostname(self) -> str | None:
        """For backwards-compatibility with urlparse. We're nice like that."""
        return self.host

    @property
    def request_uri(self) -> str:
        """Absolute path including the query string."""
        uri = self.path or "/"

        if self.query is not None:
            uri += "?" + self.query

        return uri

    @property
    def authority(self) -> str | None:
        """
        Authority component as defined in RFC 3986 3.2.
        This includes userinfo (auth), host and port.

        i.e.
            userinfo@host:port
        """
        userinfo = self.auth
        netloc = self.netloc
        if netloc is None or userinfo is None:
            return netloc
        else:
            return f"{userinfo}@{netloc}"

    @property
    def netloc(self) -> str | None:
        """
        Network location including host and port.

        If you need the equivalent of urllib.parse's ``netloc``,
        use the ``authority`` property instead.
        """
        if self.host is None:
            return None
        if self.port:
            return f"{self.host}:{self.port}"
        return self.host

    @property
    def url(self) -> str:
        """
        Convert self into a url

        This function should more or less round-trip with :func:`.parse_url`. The
        returned url may not be exactly the same as the url inputted to
        :func:`.parse_url`, but it should be equivalent by the RFC (e.g., urls
        with a blank port will have : removed).

        Example:

        .. code-block:: python

            import urllib3

            U = urllib3.util.parse_url("https://google.com/mail/")

            print(U.url)
            # "https://google.com/mail/"

            print( urllib3.util.Url("https", "username:password",
                                    "host.com", 80, "/path", "query", "fragment"
                                    ).url
                )
            # "https://username:password@host.com:80/path?query#fragment"
        """
        scheme, auth, host, port, path, query, fragment = self
        url = ""

        # We use "is not None" we want things to happen with empty strings (or 0 port)
        if scheme is not None:
            url += scheme + "://"
        if auth is not None:
            url += auth + "@"
        if host is not None:
            url += host
        if port is not None:
            url += ":" + str(port)
        if path is not None:
            url += path
        if query is not None:
            url += "?" + query
        if fragment is not None:
            url += "#" + fragment

        return url

    def __str__(self) -> str:
        return self.url


@typing.overload
def _encode_invalid_chars(
    component: str, allowed_chars: typing.Container[str]
) -> str:  # Abstract
    ...


@typing.overload
def _encode_invalid_chars(
    component: None, allowed_chars: typing.Container[str]
) -> None:  # Abstract
    ...


def _encode_invalid_chars(
    component: str | None, allowed_chars: typing.Container[str]
) -> str | None:
    """Percent-encodes a URI component without reapplying
    onto an already percent-encoded component.
    """
    if component is None:
        return component

    component = to_str(component)

    # Normalize existing percent-encoded bytes.
    # Try to see if the component we're encoding is already percent-encoded
    # so we can skip all '%' characters but still encode all others.
    component, percent_encodings = _PERCENT_RE.subn(
        lambda match: match.group(0).upper(), component
    )

    uri_bytes = component.encode("utf-8", "surrogatepass")
    is_percent_encoded = percent_encodings == uri_bytes.count(b"%")
    encoded_component = bytearray()

    for i in range(0, len(uri_bytes)):
        # Will return a single character bytestring
        byte = uri_bytes[i : i + 1]
        byte_ord = ord(byte)
        if (is_percent_encoded and byte == b"%") or (
            byte_ord < 128 and byte.decode() in allowed_chars
        ):
            encoded_component += byte
            continue
        encoded_component.extend(b"%" + (hex(byte_ord)[2:].encode().zfill(2).upper()))

    return encoded_component.decode()


def _remove_path_dot_segments(path: str) -> str:
    # See http://tools.ietf.org/html/rfc3986#section-5.2.4 for pseudo-code
    segments = path.split("/")  # Turn the path into a list of segments
    output = []  # Initialize the variable to use to store output

    for segment in segments:
        # '.' is the current directory, so ignore it, it is superfluous
        if segment == ".":
            continue
        # Anything other than '..', should be appended to the output
        if segment != "..":
            output.append(segment)
        # In this case segment == '..', if we can, we should pop the last
        # element
        elif output:
            output.pop()

    # If the path starts with '/' and the output is empty or the first string
    # is non-empty
    if path.startswith("/") and (not output or output[0]):
        output.insert(0, "")

    # If the path starts with '/.' or '/..' ensure we add one more empty
    # string to add a trailing '/'
    if path.endswith(("/.", "/..")):
        output.append("")

    return "/".join(output)


@typing.overload
def _normalize_host(host: None, scheme: str | None) -> None: ...


@typing.overload
def _normalize_host(host: str, scheme: str | None) -> str: ...


def _normalize_host(host: str | None, scheme: str | None) -> str | None:
    if host:
        if scheme in _NORMALIZABLE_SCHEMES:
            is_ipv6 = _IPV6_ADDRZ_RE.match(host)
            if is_ipv6:
                # IPv6 hosts of the form 'a::b%zone' are encoded in a URL as
                # such per RFC 6874: 'a::b%25zone'. Unquote the ZoneID
                # separator as necessary to return a valid RFC 4007 scoped IP.
                match = _ZONE_ID_RE.search(host)
                if match:
                    start, end = match.span(1)
                    zone_id = host[start:end]

                    if zone_id.startswith("%25") and zone_id != "%25":
                        zone_id = zone_id[3:]
                    else:
                        zone_id = zone_id[1:]
                    zone_id = _encode_invalid_chars(zone_id, _UNRESERVED_CHARS)
                    return f"{host[:start].lower()}%{zone_id}{host[end:]}"
                else:
                    return host.lower()
            elif not _IPV4_RE.match(host):
                return to_str(
                    b".".join([_idna_encode(label) for label in host.split(".")]),
                    "ascii",
                )
    return host


def _idna_encode(name: str) -> bytes:
    if not name.isascii():
        try:
            import idna
        except ImportError:
            raise LocationParseError(
                "Unable to parse URL without the 'idna' module"
            ) from None

        try:
            return idna.encode(name.lower(), strict=True, std3_rules=True)
        except idna.IDNAError:
            raise LocationParseError(
                f"Name '{name}' is not a valid IDNA label"
            ) from None

    return name.lower().encode("ascii")


def _encode_target(target: str) -> str:
    """Percent-encodes a request target so that there are no invalid characters

    Pre-condition for this function is that 'target' must start with '/'.
    If that is the case then _TARGET_RE will always produce a match.
    """
    match = _TARGET_RE.match(target)
    if not match:  # Defensive:
        raise LocationParseError(f"{target!r} is not a valid request URI")

    path, query = match.groups()
    encoded_target = _encode_invalid_chars(path, _PATH_CHARS)
    if query is not None:
        query = _encode_invalid_chars(query, _QUERY_CHARS)
        encoded_target += "?" + query
    return encoded_target


def parse_url(url: str) -> Url:
    """
    Given a url, return a parsed :class:`.Url` namedtuple. Best-effort is
    performed to parse incomplete urls. Fields not provided will be None.
    This parser is RFC 3986 and RFC 6874 compliant.

    The parser logic and helper functions are based heavily on
    work done in the ``rfc3986`` module.

    :param str url: URL to parse into a :class:`.Url` namedtuple.

    Partly backwards-compatible with :mod:`urllib.parse`.

    Example:

    .. code-block:: python

        import urllib3

        print( urllib3.util.parse_url('http://google.com/mail/'))
        # Url(scheme='http', host='google.com', port=None, path='/mail/', ...)

        print( urllib3.util.parse_url('google.com:80'))
        # Url(scheme=None, host='google.com', port=80, path=None, ...)

        print( urllib3.util.parse_url('/foo?bar'))
        # Url(scheme=None, host=None, port=None, path='/foo', query='bar', ...)
    """
    if not url:
        # Empty
        return Url()

    source_url = url
    if not _SCHEME_RE.search(url):
        url = "//" + url

    scheme: str | None
    authority: str | None
    auth: str | None
    host: str | None
    port: str | None
    port_int: int | None
    path: str | None
    query: str | None
    fragment: str | None

    try:
        scheme, authority, path, query, fragment = _URI_RE.match(url).groups()  # type: ignore[union-attr]
        normalize_uri = scheme is None or scheme.lower() in _NORMALIZABLE_SCHEMES

        if scheme:
            scheme = scheme.lower()

        if authority:
            auth, _, host_port = authority.rpartition("@")
            auth = auth or None
            host, port = _HOST_PORT_RE.match(host_port).groups()  # type: ignore[union-attr]
            if auth and normalize_uri:
                auth = _encode_invalid_chars(auth, _USERINFO_CHARS)
            if port == "":
                port = None
        else:
            auth, host, port = None, None, None

        if port is not None:
            port_int = int(port)
            if not (0 <= port_int <= 65535):
                raise LocationParseError(url)
        else:
            port_int = None

        host = _normalize_host(host, scheme)

        if normalize_uri and path:
            path = _remove_path_dot_segments(path)
            path = _encode_invalid_chars(path, _PATH_CHARS)
        if normalize_uri and query:
            query = _encode_invalid_chars(query, _QUERY_CHARS)
        if normalize_uri and fragment:
            fragment = _encode_invalid_chars(fragment, _FRAGMENT_CHARS)

    except (ValueError, AttributeError) as e:
        raise LocationParseError(source_url) from e

    # For the sake of backwards compatibility we put empty
    # string values for path if there are any defined values
    # beyond the path in the URL.
    # TODO: Remove this when we break backwards compatibility.
    if not path:
        if query is not None or fragment is not None:
            path = ""
        else:
            path = None

    return Url(
        scheme=scheme,
        auth=auth,
        host=host,
        port=port_int,
        path=path,
        query=query,
        fragment=fragment,
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\attr\_funcs.py ===
# SPDX-License-Identifier: MIT


import copy

from ._compat import PY_3_9_PLUS, get_generic_base
from ._make import _OBJ_SETATTR, NOTHING, fields
from .exceptions import AttrsAttributeNotFoundError


def asdict(
    inst,
    recurse=True,
    filter=None,
    dict_factory=dict,
    retain_collection_types=False,
    value_serializer=None,
):
    """
    Return the *attrs* attribute values of *inst* as a dict.

    Optionally recurse into other *attrs*-decorated classes.

    Args:
        inst: Instance of an *attrs*-decorated class.

        recurse (bool): Recurse into classes that are also *attrs*-decorated.

        filter (~typing.Callable):
            A callable whose return code determines whether an attribute or
            element is included (`True`) or dropped (`False`).  Is called with
            the `attrs.Attribute` as the first argument and the value as the
            second argument.

        dict_factory (~typing.Callable):
            A callable to produce dictionaries from.  For example, to produce
            ordered dictionaries instead of normal Python dictionaries, pass in
            ``collections.OrderedDict``.

        retain_collection_types (bool):
            Do not convert to `list` when encountering an attribute whose type
            is `tuple` or `set`.  Only meaningful if *recurse* is `True`.

        value_serializer (typing.Callable | None):
            A hook that is called for every attribute or dict key/value.  It
            receives the current instance, field and value and must return the
            (updated) value.  The hook is run *after* the optional *filter* has
            been applied.

    Returns:
        Return type of *dict_factory*.

    Raises:
        attrs.exceptions.NotAnAttrsClassError:
            If *cls* is not an *attrs* class.

    ..  versionadded:: 16.0.0 *dict_factory*
    ..  versionadded:: 16.1.0 *retain_collection_types*
    ..  versionadded:: 20.3.0 *value_serializer*
    ..  versionadded:: 21.3.0
        If a dict has a collection for a key, it is serialized as a tuple.
    """
    attrs = fields(inst.__class__)
    rv = dict_factory()
    for a in attrs:
        v = getattr(inst, a.name)
        if filter is not None and not filter(a, v):
            continue

        if value_serializer is not None:
            v = value_serializer(inst, a, v)

        if recurse is True:
            if has(v.__class__):
                rv[a.name] = asdict(
                    v,
                    recurse=True,
                    filter=filter,
                    dict_factory=dict_factory,
                    retain_collection_types=retain_collection_types,
                    value_serializer=value_serializer,
                )
            elif isinstance(v, (tuple, list, set, frozenset)):
                cf = v.__class__ if retain_collection_types is True else list
                items = [
                    _asdict_anything(
                        i,
                        is_key=False,
                        filter=filter,
                        dict_factory=dict_factory,
                        retain_collection_types=retain_collection_types,
                        value_serializer=value_serializer,
                    )
                    for i in v
                ]
                try:
                    rv[a.name] = cf(items)
                except TypeError:
                    if not issubclass(cf, tuple):
                        raise
                    # Workaround for TypeError: cf.__new__() missing 1 required
                    # positional argument (which appears, for a namedturle)
                    rv[a.name] = cf(*items)
            elif isinstance(v, dict):
                df = dict_factory
                rv[a.name] = df(
                    (
                        _asdict_anything(
                            kk,
                            is_key=True,
                            filter=filter,
                            dict_factory=df,
                            retain_collection_types=retain_collection_types,
                            value_serializer=value_serializer,
                        ),
                        _asdict_anything(
                            vv,
                            is_key=False,
                            filter=filter,
                            dict_factory=df,
                            retain_collection_types=retain_collection_types,
                            value_serializer=value_serializer,
                        ),
                    )
                    for kk, vv in v.items()
                )
            else:
                rv[a.name] = v
        else:
            rv[a.name] = v
    return rv


def _asdict_anything(
    val,
    is_key,
    filter,
    dict_factory,
    retain_collection_types,
    value_serializer,
):
    """
    ``asdict`` only works on attrs instances, this works on anything.
    """
    if getattr(val.__class__, "__attrs_attrs__", None) is not None:
        # Attrs class.
        rv = asdict(
            val,
            recurse=True,
            filter=filter,
            dict_factory=dict_factory,
            retain_collection_types=retain_collection_types,
            value_serializer=value_serializer,
        )
    elif isinstance(val, (tuple, list, set, frozenset)):
        if retain_collection_types is True:
            cf = val.__class__
        elif is_key:
            cf = tuple
        else:
            cf = list

        rv = cf(
            [
                _asdict_anything(
                    i,
                    is_key=False,
                    filter=filter,
                    dict_factory=dict_factory,
                    retain_collection_types=retain_collection_types,
                    value_serializer=value_serializer,
                )
                for i in val
            ]
        )
    elif isinstance(val, dict):
        df = dict_factory
        rv = df(
            (
                _asdict_anything(
                    kk,
                    is_key=True,
                    filter=filter,
                    dict_factory=df,
                    retain_collection_types=retain_collection_types,
                    value_serializer=value_serializer,
                ),
                _asdict_anything(
                    vv,
                    is_key=False,
                    filter=filter,
                    dict_factory=df,
                    retain_collection_types=retain_collection_types,
                    value_serializer=value_serializer,
                ),
            )
            for kk, vv in val.items()
        )
    else:
        rv = val
        if value_serializer is not None:
            rv = value_serializer(None, None, rv)

    return rv


def astuple(
    inst,
    recurse=True,
    filter=None,
    tuple_factory=tuple,
    retain_collection_types=False,
):
    """
    Return the *attrs* attribute values of *inst* as a tuple.

    Optionally recurse into other *attrs*-decorated classes.

    Args:
        inst: Instance of an *attrs*-decorated class.

        recurse (bool):
            Recurse into classes that are also *attrs*-decorated.

        filter (~typing.Callable):
            A callable whose return code determines whether an attribute or
            element is included (`True`) or dropped (`False`).  Is called with
            the `attrs.Attribute` as the first argument and the value as the
            second argument.

        tuple_factory (~typing.Callable):
            A callable to produce tuples from. For example, to produce lists
            instead of tuples.

        retain_collection_types (bool):
            Do not convert to `list` or `dict` when encountering an attribute
            which type is `tuple`, `dict` or `set`. Only meaningful if
            *recurse* is `True`.

    Returns:
        Return type of *tuple_factory*

    Raises:
        attrs.exceptions.NotAnAttrsClassError:
            If *cls* is not an *attrs* class.

    ..  versionadded:: 16.2.0
    """
    attrs = fields(inst.__class__)
    rv = []
    retain = retain_collection_types  # Very long. :/
    for a in attrs:
        v = getattr(inst, a.name)
        if filter is not None and not filter(a, v):
            continue
        if recurse is True:
            if has(v.__class__):
                rv.append(
                    astuple(
                        v,
                        recurse=True,
                        filter=filter,
                        tuple_factory=tuple_factory,
                        retain_collection_types=retain,
                    )
                )
            elif isinstance(v, (tuple, list, set, frozenset)):
                cf = v.__class__ if retain is True else list
                items = [
                    (
                        astuple(
                            j,
                            recurse=True,
                            filter=filter,
                            tuple_factory=tuple_factory,
                            retain_collection_types=retain,
                        )
                        if has(j.__class__)
                        else j
                    )
                    for j in v
                ]
                try:
                    rv.append(cf(items))
                except TypeError:
                    if not issubclass(cf, tuple):
                        raise
                    # Workaround for TypeError: cf.__new__() missing 1 required
                    # positional argument (which appears, for a namedturle)
                    rv.append(cf(*items))
            elif isinstance(v, dict):
                df = v.__class__ if retain is True else dict
                rv.append(
                    df(
                        (
                            (
                                astuple(
                                    kk,
                                    tuple_factory=tuple_factory,
                                    retain_collection_types=retain,
                                )
                                if has(kk.__class__)
                                else kk
                            ),
                            (
                                astuple(
                                    vv,
                                    tuple_factory=tuple_factory,
                                    retain_collection_types=retain,
                                )
                                if has(vv.__class__)
                                else vv
                            ),
                        )
                        for kk, vv in v.items()
                    )
                )
            else:
                rv.append(v)
        else:
            rv.append(v)

    return rv if tuple_factory is list else tuple_factory(rv)


def has(cls):
    """
    Check whether *cls* is a class with *attrs* attributes.

    Args:
        cls (type): Class to introspect.

    Raises:
        TypeError: If *cls* is not a class.

    Returns:
        bool:
    """
    attrs = getattr(cls, "__attrs_attrs__", None)
    if attrs is not None:
        return True

    # No attrs, maybe it's a specialized generic (A[str])?
    generic_base = get_generic_base(cls)
    if generic_base is not None:
        generic_attrs = getattr(generic_base, "__attrs_attrs__", None)
        if generic_attrs is not None:
            # Stick it on here for speed next time.
            cls.__attrs_attrs__ = generic_attrs
        return generic_attrs is not None
    return False


def assoc(inst, **changes):
    """
    Copy *inst* and apply *changes*.

    This is different from `evolve` that applies the changes to the arguments
    that create the new instance.

    `evolve`'s behavior is preferable, but there are `edge cases`_ where it
    doesn't work. Therefore `assoc` is deprecated, but will not be removed.

    .. _`edge cases`: https://github.com/python-attrs/attrs/issues/251

    Args:
        inst: Instance of a class with *attrs* attributes.

        changes: Keyword changes in the new copy.

    Returns:
        A copy of inst with *changes* incorporated.

    Raises:
        attrs.exceptions.AttrsAttributeNotFoundError:
            If *attr_name* couldn't be found on *cls*.

        attrs.exceptions.NotAnAttrsClassError:
            If *cls* is not an *attrs* class.

    ..  deprecated:: 17.1.0
        Use `attrs.evolve` instead if you can. This function will not be
        removed du to the slightly different approach compared to
        `attrs.evolve`, though.
    """
    new = copy.copy(inst)
    attrs = fields(inst.__class__)
    for k, v in changes.items():
        a = getattr(attrs, k, NOTHING)
        if a is NOTHING:
            msg = f"{k} is not an attrs attribute on {new.__class__}."
            raise AttrsAttributeNotFoundError(msg)
        _OBJ_SETATTR(new, k, v)
    return new


def resolve_types(
    cls, globalns=None, localns=None, attribs=None, include_extras=True
):
    """
    Resolve any strings and forward annotations in type annotations.

    This is only required if you need concrete types in :class:`Attribute`'s
    *type* field. In other words, you don't need to resolve your types if you
    only use them for static type checking.

    With no arguments, names will be looked up in the module in which the class
    was created. If this is not what you want, for example, if the name only
    exists inside a method, you may pass *globalns* or *localns* to specify
    other dictionaries in which to look up these names. See the docs of
    `typing.get_type_hints` for more details.

    Args:
        cls (type): Class to resolve.

        globalns (dict | None): Dictionary containing global variables.

        localns (dict | None): Dictionary containing local variables.

        attribs (list | None):
            List of attribs for the given class. This is necessary when calling
            from inside a ``field_transformer`` since *cls* is not an *attrs*
            class yet.

        include_extras (bool):
            Resolve more accurately, if possible. Pass ``include_extras`` to
            ``typing.get_hints``, if supported by the typing module. On
            supported Python versions (3.9+), this resolves the types more
            accurately.

    Raises:
        TypeError: If *cls* is not a class.

        attrs.exceptions.NotAnAttrsClassError:
            If *cls* is not an *attrs* class and you didn't pass any attribs.

        NameError: If types cannot be resolved because of missing variables.

    Returns:
        *cls* so you can use this function also as a class decorator. Please
        note that you have to apply it **after** `attrs.define`. That means the
        decorator has to come in the line **before** `attrs.define`.

    ..  versionadded:: 20.1.0
    ..  versionadded:: 21.1.0 *attribs*
    ..  versionadded:: 23.1.0 *include_extras*
    """
    # Since calling get_type_hints is expensive we cache whether we've
    # done it already.
    if getattr(cls, "__attrs_types_resolved__", None) != cls:
        import typing

        kwargs = {"globalns": globalns, "localns": localns}

        if PY_3_9_PLUS:
            kwargs["include_extras"] = include_extras

        hints = typing.get_type_hints(cls, **kwargs)
        for field in fields(cls) if attribs is None else attribs:
            if field.name in hints:
                # Since fields have been frozen we must work around it.
                _OBJ_SETATTR(field, "type", hints[field.name])
        # We store the class we resolved so that subclasses know they haven't
        # been resolved.
        cls.__attrs_types_resolved__ = cls

    # Return the class so you can use it as a decorator too.
    return cls

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\mpmath\__init__.py ===
__version__ = '1.3.0'

from .usertools import monitor, timing

from .ctx_fp import FPContext
from .ctx_mp import MPContext
from .ctx_iv import MPIntervalContext

fp = FPContext()
mp = MPContext()
iv = MPIntervalContext()

fp._mp = mp
mp._mp = mp
iv._mp = mp
mp._fp = fp
fp._fp = fp
mp._iv = iv
fp._iv = iv
iv._iv = iv

# XXX: extremely bad pickle hack
from . import ctx_mp as _ctx_mp
_ctx_mp._mpf_module.mpf = mp.mpf
_ctx_mp._mpf_module.mpc = mp.mpc

make_mpf = mp.make_mpf
make_mpc = mp.make_mpc

extraprec = mp.extraprec
extradps = mp.extradps
workprec = mp.workprec
workdps = mp.workdps
autoprec = mp.autoprec
maxcalls = mp.maxcalls
memoize = mp.memoize

mag = mp.mag

bernfrac = mp.bernfrac

qfrom = mp.qfrom
mfrom = mp.mfrom
kfrom = mp.kfrom
taufrom = mp.taufrom
qbarfrom = mp.qbarfrom
ellipfun = mp.ellipfun
jtheta = mp.jtheta
kleinj = mp.kleinj
eta = mp.eta

qp = mp.qp
qhyper = mp.qhyper
qgamma = mp.qgamma
qfac = mp.qfac

nint_distance = mp.nint_distance

plot = mp.plot
cplot = mp.cplot
splot = mp.splot

odefun = mp.odefun

jacobian = mp.jacobian
findroot = mp.findroot
multiplicity = mp.multiplicity

isinf = mp.isinf
isnan = mp.isnan
isnormal = mp.isnormal
isint = mp.isint
isfinite = mp.isfinite
almosteq = mp.almosteq
nan = mp.nan
rand = mp.rand

absmin = mp.absmin
absmax = mp.absmax

fraction = mp.fraction

linspace = mp.linspace
arange = mp.arange

mpmathify = convert = mp.convert
mpc = mp.mpc

mpi = iv._mpi

nstr = mp.nstr
nprint = mp.nprint
chop = mp.chop

fneg = mp.fneg
fadd = mp.fadd
fsub = mp.fsub
fmul = mp.fmul
fdiv = mp.fdiv
fprod = mp.fprod

quad = mp.quad
quadgl = mp.quadgl
quadts = mp.quadts
quadosc = mp.quadosc
quadsubdiv = mp.quadsubdiv

invertlaplace = mp.invertlaplace
invlaptalbot = mp.invlaptalbot
invlapstehfest = mp.invlapstehfest
invlapdehoog = mp.invlapdehoog

pslq = mp.pslq
identify = mp.identify
findpoly = mp.findpoly

richardson = mp.richardson
shanks = mp.shanks
levin = mp.levin
cohen_alt = mp.cohen_alt
nsum = mp.nsum
nprod = mp.nprod
difference = mp.difference
diff = mp.diff
diffs = mp.diffs
diffs_prod = mp.diffs_prod
diffs_exp = mp.diffs_exp
diffun = mp.diffun
differint = mp.differint
taylor = mp.taylor
pade = mp.pade
polyval = mp.polyval
polyroots = mp.polyroots
fourier = mp.fourier
fourierval = mp.fourierval
sumem = mp.sumem
sumap = mp.sumap
chebyfit = mp.chebyfit
limit = mp.limit

matrix = mp.matrix
eye = mp.eye
diag = mp.diag
zeros = mp.zeros
ones = mp.ones
hilbert = mp.hilbert
randmatrix = mp.randmatrix
swap_row = mp.swap_row
extend = mp.extend
norm = mp.norm
mnorm = mp.mnorm

lu_solve = mp.lu_solve
lu = mp.lu
qr = mp.qr
unitvector = mp.unitvector
inverse = mp.inverse
residual = mp.residual
qr_solve = mp.qr_solve
cholesky = mp.cholesky
cholesky_solve = mp.cholesky_solve
det = mp.det
cond = mp.cond
hessenberg = mp.hessenberg
schur = mp.schur
eig = mp.eig
eig_sort = mp.eig_sort
eigsy = mp.eigsy
eighe = mp.eighe
eigh = mp.eigh
svd_r = mp.svd_r
svd_c = mp.svd_c
svd = mp.svd
gauss_quadrature = mp.gauss_quadrature

expm = mp.expm
sqrtm = mp.sqrtm
powm = mp.powm
logm = mp.logm
sinm = mp.sinm
cosm = mp.cosm

mpf = mp.mpf
j = mp.j
exp = mp.exp
expj = mp.expj
expjpi = mp.expjpi
ln = mp.ln
im = mp.im
re = mp.re
inf = mp.inf
ninf = mp.ninf
sign = mp.sign

eps = mp.eps
pi = mp.pi
ln2 = mp.ln2
ln10 = mp.ln10
phi = mp.phi
e = mp.e
euler = mp.euler
catalan = mp.catalan
khinchin = mp.khinchin
glaisher = mp.glaisher
apery = mp.apery
degree = mp.degree
twinprime = mp.twinprime
mertens = mp.mertens

ldexp = mp.ldexp
frexp = mp.frexp

fsum = mp.fsum
fdot = mp.fdot

sqrt = mp.sqrt
cbrt = mp.cbrt
exp = mp.exp
ln = mp.ln
log = mp.log
log10 = mp.log10
power = mp.power
cos = mp.cos
sin = mp.sin
tan = mp.tan
cosh = mp.cosh
sinh = mp.sinh
tanh = mp.tanh
acos = mp.acos
asin = mp.asin
atan = mp.atan
asinh = mp.asinh
acosh = mp.acosh
atanh = mp.atanh
sec = mp.sec
csc = mp.csc
cot = mp.cot
sech = mp.sech
csch = mp.csch
coth = mp.coth
asec = mp.asec
acsc = mp.acsc
acot = mp.acot
asech = mp.asech
acsch = mp.acsch
acoth = mp.acoth
cospi = mp.cospi
sinpi = mp.sinpi
sinc = mp.sinc
sincpi = mp.sincpi
cos_sin = mp.cos_sin
cospi_sinpi = mp.cospi_sinpi
fabs = mp.fabs
re = mp.re
im = mp.im
conj = mp.conj
floor = mp.floor
ceil = mp.ceil
nint = mp.nint
frac = mp.frac
root = mp.root
nthroot = mp.nthroot
hypot = mp.hypot
fmod = mp.fmod
ldexp = mp.ldexp
frexp = mp.frexp
sign = mp.sign
arg = mp.arg
phase = mp.phase
polar = mp.polar
rect = mp.rect
degrees = mp.degrees
radians = mp.radians
atan2 = mp.atan2
fib = mp.fib
fibonacci = mp.fibonacci
lambertw = mp.lambertw
zeta = mp.zeta
altzeta = mp.altzeta
gamma = mp.gamma
rgamma = mp.rgamma
factorial = mp.factorial
fac = mp.fac
fac2 = mp.fac2
beta = mp.beta
betainc = mp.betainc
psi = mp.psi
#psi0 = mp.psi0
#psi1 = mp.psi1
#psi2 = mp.psi2
#psi3 = mp.psi3
polygamma = mp.polygamma
digamma = mp.digamma
#trigamma = mp.trigamma
#tetragamma = mp.tetragamma
#pentagamma = mp.pentagamma
harmonic = mp.harmonic
bernoulli = mp.bernoulli
bernfrac = mp.bernfrac
stieltjes = mp.stieltjes
hurwitz = mp.hurwitz
dirichlet = mp.dirichlet
bernpoly = mp.bernpoly
eulerpoly = mp.eulerpoly
eulernum = mp.eulernum
polylog = mp.polylog
clsin = mp.clsin
clcos = mp.clcos
gammainc = mp.gammainc
gammaprod = mp.gammaprod
binomial = mp.binomial
rf = mp.rf
ff = mp.ff
hyper = mp.hyper
hyp0f1 = mp.hyp0f1
hyp1f1 = mp.hyp1f1
hyp1f2 = mp.hyp1f2
hyp2f1 = mp.hyp2f1
hyp2f2 = mp.hyp2f2
hyp2f0 = mp.hyp2f0
hyp2f3 = mp.hyp2f3
hyp3f2 = mp.hyp3f2
hyperu = mp.hyperu
hypercomb = mp.hypercomb
meijerg = mp.meijerg
appellf1 = mp.appellf1
appellf2 = mp.appellf2
appellf3 = mp.appellf3
appellf4 = mp.appellf4
hyper2d = mp.hyper2d
bihyper = mp.bihyper
erf = mp.erf
erfc = mp.erfc
erfi = mp.erfi
erfinv = mp.erfinv
npdf = mp.npdf
ncdf = mp.ncdf
expint = mp.expint
e1 = mp.e1
ei = mp.ei
li = mp.li
ci = mp.ci
si = mp.si
chi = mp.chi
shi = mp.shi
fresnels = mp.fresnels
fresnelc = mp.fresnelc
airyai = mp.airyai
airybi = mp.airybi
airyaizero = mp.airyaizero
airybizero = mp.airybizero
scorergi = mp.scorergi
scorerhi = mp.scorerhi
ellipk = mp.ellipk
ellipe = mp.ellipe
ellipf = mp.ellipf
ellippi = mp.ellippi
elliprc = mp.elliprc
elliprj = mp.elliprj
elliprf = mp.elliprf
elliprd = mp.elliprd
elliprg = mp.elliprg
agm = mp.agm
jacobi = mp.jacobi
chebyt = mp.chebyt
chebyu = mp.chebyu
legendre = mp.legendre
legenp = mp.legenp
legenq = mp.legenq
hermite = mp.hermite
pcfd = mp.pcfd
pcfu = mp.pcfu
pcfv = mp.pcfv
pcfw = mp.pcfw
gegenbauer = mp.gegenbauer
laguerre = mp.laguerre
spherharm = mp.spherharm
besselj = mp.besselj
j0 = mp.j0
j1 = mp.j1
besseli = mp.besseli
bessely = mp.bessely
besselk = mp.besselk
besseljzero = mp.besseljzero
besselyzero = mp.besselyzero
hankel1 = mp.hankel1
hankel2 = mp.hankel2
struveh = mp.struveh
struvel = mp.struvel
angerj = mp.angerj
webere = mp.webere
lommels1 = mp.lommels1
lommels2 = mp.lommels2
whitm = mp.whitm
whitw = mp.whitw
ber = mp.ber
bei = mp.bei
ker = mp.ker
kei = mp.kei
coulombc = mp.coulombc
coulombf = mp.coulombf
coulombg = mp.coulombg
barnesg = mp.barnesg
superfac = mp.superfac
hyperfac = mp.hyperfac
loggamma = mp.loggamma
siegeltheta = mp.siegeltheta
siegelz = mp.siegelz
grampoint = mp.grampoint
zetazero = mp.zetazero
riemannr = mp.riemannr
primepi = mp.primepi
primepi2 = mp.primepi2
primezeta = mp.primezeta
bell = mp.bell
polyexp = mp.polyexp
expm1 = mp.expm1
log1p = mp.log1p
powm1 = mp.powm1
unitroots = mp.unitroots
cyclotomic = mp.cyclotomic
mangoldt = mp.mangoldt
secondzeta = mp.secondzeta
nzeros = mp.nzeros
backlunds = mp.backlunds
lerchphi = mp.lerchphi
stirling1 = mp.stirling1
stirling2 = mp.stirling2
squarew = mp.squarew
trianglew = mp.trianglew
sawtoothw = mp.sawtoothw
unit_triangle = mp.unit_triangle
sigmoid = mp.sigmoid

# be careful when changing this name, don't use test*!
def runtests():
    """
    Run all mpmath tests and print output.
    """
    import os.path
    from inspect import getsourcefile
    from .tests import runtests as tests
    testdir = os.path.dirname(os.path.abspath(getsourcefile(tests)))
    importdir = os.path.abspath(testdir + '/../..')
    tests.testit(importdir, testdir)

def doctests(filter=[]):
    import sys
    from timeit import default_timer as clock
    for i, arg in enumerate(sys.argv):
        if '__init__.py' in arg:
            filter = [sn for sn in sys.argv[i+1:] if not sn.startswith("-")]
            break
    import doctest
    globs = globals().copy()
    for obj in globs: #sorted(globs.keys()):
        if filter:
            if not sum([pat in obj for pat in filter]):
                continue
        sys.stdout.write(str(obj) + " ")
        sys.stdout.flush()
        t1 = clock()
        doctest.run_docstring_examples(globs[obj], {}, verbose=("-v" in sys.argv))
        t2 = clock()
        print(round(t2-t1, 3))

if __name__ == '__main__':
    doctests()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\sql\_typing.py ===
# sql/_typing.py
# Copyright (C) 2022-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

from __future__ import annotations

import operator
from typing import Any
from typing import Callable
from typing import Dict
from typing import Generic
from typing import Iterable
from typing import Mapping
from typing import NoReturn
from typing import Optional
from typing import overload
from typing import Set
from typing import Tuple
from typing import Type
from typing import TYPE_CHECKING
from typing import TypeVar
from typing import Union

from . import roles
from .. import exc
from .. import util
from ..inspection import Inspectable
from ..util.typing import Literal
from ..util.typing import Protocol
from ..util.typing import TypeAlias

if TYPE_CHECKING:
    from datetime import date
    from datetime import datetime
    from datetime import time
    from datetime import timedelta
    from decimal import Decimal
    from uuid import UUID

    from .base import Executable
    from .compiler import Compiled
    from .compiler import DDLCompiler
    from .compiler import SQLCompiler
    from .dml import UpdateBase
    from .dml import ValuesBase
    from .elements import ClauseElement
    from .elements import ColumnElement
    from .elements import KeyedColumnElement
    from .elements import quoted_name
    from .elements import SQLCoreOperations
    from .elements import TextClause
    from .lambdas import LambdaElement
    from .roles import FromClauseRole
    from .schema import Column
    from .selectable import Alias
    from .selectable import CompoundSelect
    from .selectable import CTE
    from .selectable import FromClause
    from .selectable import Join
    from .selectable import NamedFromClause
    from .selectable import ReturnsRows
    from .selectable import Select
    from .selectable import Selectable
    from .selectable import SelectBase
    from .selectable import Subquery
    from .selectable import TableClause
    from .sqltypes import TableValueType
    from .sqltypes import TupleType
    from .type_api import TypeEngine
    from ..engine import Connection
    from ..engine import Dialect
    from ..engine import Engine
    from ..engine.mock import MockConnection
    from ..util.typing import TypeGuard

_T = TypeVar("_T", bound=Any)
_T_co = TypeVar("_T_co", bound=Any, covariant=True)


_CE = TypeVar("_CE", bound="ColumnElement[Any]")

_CLE = TypeVar("_CLE", bound="ClauseElement")


class _HasClauseElement(Protocol, Generic[_T_co]):
    """indicates a class that has a __clause_element__() method"""

    def __clause_element__(self) -> roles.ExpressionElementRole[_T_co]: ...


class _CoreAdapterProto(Protocol):
    """protocol for the ClauseAdapter/ColumnAdapter.traverse() method."""

    def __call__(self, obj: _CE) -> _CE: ...


class _HasDialect(Protocol):
    """protocol for Engine/Connection-like objects that have dialect
    attribute.
    """

    @property
    def dialect(self) -> Dialect: ...


# match column types that are not ORM entities
_NOT_ENTITY = TypeVar(
    "_NOT_ENTITY",
    int,
    str,
    bool,
    "datetime",
    "date",
    "time",
    "timedelta",
    "UUID",
    float,
    "Decimal",
)

_StarOrOne = Literal["*", 1]

_MAYBE_ENTITY = TypeVar(
    "_MAYBE_ENTITY",
    roles.ColumnsClauseRole,
    _StarOrOne,
    Type[Any],
    Inspectable[_HasClauseElement[Any]],
    _HasClauseElement[Any],
)


# convention:
# XYZArgument - something that the end user is passing to a public API method
# XYZElement - the internal representation that we use for the thing.
# the coercions system is responsible for converting from XYZArgument to
# XYZElement.

_TextCoercedExpressionArgument = Union[
    str,
    "TextClause",
    "ColumnElement[_T]",
    _HasClauseElement[_T],
    roles.ExpressionElementRole[_T],
]

_ColumnsClauseArgument = Union[
    roles.TypedColumnsClauseRole[_T],
    roles.ColumnsClauseRole,
    "SQLCoreOperations[_T]",
    _StarOrOne,
    Type[_T],
    Inspectable[_HasClauseElement[_T]],
    _HasClauseElement[_T],
]
"""open-ended SELECT columns clause argument.

Includes column expressions, tables, ORM mapped entities, a few literal values.

This type is used for lists of columns  / entities to be returned in result
sets; select(...), insert().returning(...), etc.


"""

_TypedColumnClauseArgument = Union[
    roles.TypedColumnsClauseRole[_T],
    "SQLCoreOperations[_T]",
    Type[_T],
]

_TP = TypeVar("_TP", bound=Tuple[Any, ...])

_T0 = TypeVar("_T0", bound=Any)
_T1 = TypeVar("_T1", bound=Any)
_T2 = TypeVar("_T2", bound=Any)
_T3 = TypeVar("_T3", bound=Any)
_T4 = TypeVar("_T4", bound=Any)
_T5 = TypeVar("_T5", bound=Any)
_T6 = TypeVar("_T6", bound=Any)
_T7 = TypeVar("_T7", bound=Any)
_T8 = TypeVar("_T8", bound=Any)
_T9 = TypeVar("_T9", bound=Any)


_ColumnExpressionArgument = Union[
    "ColumnElement[_T]",
    _HasClauseElement[_T],
    "SQLCoreOperations[_T]",
    roles.ExpressionElementRole[_T],
    roles.TypedColumnsClauseRole[_T],
    Callable[[], "ColumnElement[_T]"],
    "LambdaElement",
]
"See docs in public alias ColumnExpressionArgument."

ColumnExpressionArgument: TypeAlias = _ColumnExpressionArgument[_T]
"""Narrower "column expression" argument.

This type is used for all the other "column" kinds of expressions that
typically represent a single SQL column expression, not a set of columns the
way a table or ORM entity does.

This includes ColumnElement, or ORM-mapped attributes that will have a
``__clause_element__()`` method, it also has the ExpressionElementRole
overall which brings in the TextClause object also.

.. versionadded:: 2.0.13

"""

_ColumnExpressionOrLiteralArgument = Union[Any, _ColumnExpressionArgument[_T]]

_ColumnExpressionOrStrLabelArgument = Union[str, _ColumnExpressionArgument[_T]]

_ByArgument = Union[
    Iterable[_ColumnExpressionOrStrLabelArgument[Any]],
    _ColumnExpressionOrStrLabelArgument[Any],
]
"""Used for keyword-based ``order_by`` and ``partition_by`` parameters."""


_InfoType = Dict[Any, Any]
"""the .info dictionary accepted and used throughout Core /ORM"""

_FromClauseArgument = Union[
    roles.FromClauseRole,
    Type[Any],
    Inspectable[_HasClauseElement[Any]],
    _HasClauseElement[Any],
]
"""A FROM clause, like we would send to select().select_from().

Also accommodates ORM entities and related constructs.

"""

_JoinTargetArgument = Union[_FromClauseArgument, roles.JoinTargetRole]
"""target for join() builds on _FromClauseArgument to include additional
join target roles such as those which come from the ORM.

"""

_OnClauseArgument = Union[_ColumnExpressionArgument[Any], roles.OnClauseRole]
"""target for an ON clause, includes additional roles such as those which
come from the ORM.

"""

_SelectStatementForCompoundArgument = Union[
    "Select[_TP]",
    "CompoundSelect[_TP]",
    roles.CompoundElementRole,
]
"""SELECT statement acceptable by ``union()`` and other SQL set operations"""

_DMLColumnArgument = Union[
    str,
    _HasClauseElement[Any],
    roles.DMLColumnRole,
    "SQLCoreOperations[Any]",
]
"""A DML column expression.  This is a "key" inside of insert().values(),
update().values(), and related.

These are usually strings or SQL table columns.

There's also edge cases like JSON expression assignment, which we would want
the DMLColumnRole to be able to accommodate.

"""

_DMLKey = TypeVar("_DMLKey", bound=_DMLColumnArgument)
_DMLColumnKeyMapping = Mapping[_DMLKey, Any]


_DDLColumnArgument = Union[str, "Column[Any]", roles.DDLConstraintColumnRole]
"""DDL column.

used for :class:`.PrimaryKeyConstraint`, :class:`.UniqueConstraint`, etc.

"""

_DMLTableArgument = Union[
    "TableClause",
    "Join",
    "Alias",
    "CTE",
    Type[Any],
    Inspectable[_HasClauseElement[Any]],
    _HasClauseElement[Any],
]

_PropagateAttrsType = util.immutabledict[str, Any]

_TypeEngineArgument = Union[Type["TypeEngine[_T]"], "TypeEngine[_T]"]

_EquivalentColumnMap = Dict["ColumnElement[Any]", Set["ColumnElement[Any]"]]

_LimitOffsetType = Union[int, _ColumnExpressionArgument[int], None]

_AutoIncrementType = Union[bool, Literal["auto", "ignore_fk"]]

_CreateDropBind = Union["Engine", "Connection", "MockConnection"]

if TYPE_CHECKING:

    def is_sql_compiler(c: Compiled) -> TypeGuard[SQLCompiler]: ...

    def is_ddl_compiler(c: Compiled) -> TypeGuard[DDLCompiler]: ...

    def is_named_from_clause(
        t: FromClauseRole,
    ) -> TypeGuard[NamedFromClause]: ...

    def is_column_element(
        c: ClauseElement,
    ) -> TypeGuard[ColumnElement[Any]]: ...

    def is_keyed_column_element(
        c: ClauseElement,
    ) -> TypeGuard[KeyedColumnElement[Any]]: ...

    def is_text_clause(c: ClauseElement) -> TypeGuard[TextClause]: ...

    def is_from_clause(c: ClauseElement) -> TypeGuard[FromClause]: ...

    def is_tuple_type(t: TypeEngine[Any]) -> TypeGuard[TupleType]: ...

    def is_table_value_type(
        t: TypeEngine[Any],
    ) -> TypeGuard[TableValueType]: ...

    def is_selectable(t: Any) -> TypeGuard[Selectable]: ...

    def is_select_base(
        t: Union[Executable, ReturnsRows]
    ) -> TypeGuard[SelectBase]: ...

    def is_select_statement(
        t: Union[Executable, ReturnsRows]
    ) -> TypeGuard[Select[Any]]: ...

    def is_table(t: FromClause) -> TypeGuard[TableClause]: ...

    def is_subquery(t: FromClause) -> TypeGuard[Subquery]: ...

    def is_dml(c: ClauseElement) -> TypeGuard[UpdateBase]: ...

else:
    is_sql_compiler = operator.attrgetter("is_sql")
    is_ddl_compiler = operator.attrgetter("is_ddl")
    is_named_from_clause = operator.attrgetter("named_with_column")
    is_column_element = operator.attrgetter("_is_column_element")
    is_keyed_column_element = operator.attrgetter("_is_keyed_column_element")
    is_text_clause = operator.attrgetter("_is_text_clause")
    is_from_clause = operator.attrgetter("_is_from_clause")
    is_tuple_type = operator.attrgetter("_is_tuple_type")
    is_table_value_type = operator.attrgetter("_is_table_value")
    is_selectable = operator.attrgetter("is_selectable")
    is_select_base = operator.attrgetter("_is_select_base")
    is_select_statement = operator.attrgetter("_is_select_statement")
    is_table = operator.attrgetter("_is_table")
    is_subquery = operator.attrgetter("_is_subquery")
    is_dml = operator.attrgetter("is_dml")


def has_schema_attr(t: FromClauseRole) -> TypeGuard[TableClause]:
    return hasattr(t, "schema")


def is_quoted_name(s: str) -> TypeGuard[quoted_name]:
    return hasattr(s, "quote")


def is_has_clause_element(s: object) -> TypeGuard[_HasClauseElement[Any]]:
    return hasattr(s, "__clause_element__")


def is_insert_update(c: ClauseElement) -> TypeGuard[ValuesBase]:
    return c.is_dml and (c.is_insert or c.is_update)  # type: ignore


def _no_kw() -> exc.ArgumentError:
    return exc.ArgumentError(
        "Additional keyword arguments are not accepted by this "
        "function/method.  The presence of **kw is for pep-484 typing purposes"
    )


def _unexpected_kw(methname: str, kw: Dict[str, Any]) -> NoReturn:
    k = list(kw)[0]
    raise TypeError(f"{methname} got an unexpected keyword argument '{k}'")


@overload
def Nullable(
    val: "SQLCoreOperations[_T]",
) -> "SQLCoreOperations[Optional[_T]]": ...


@overload
def Nullable(
    val: roles.ExpressionElementRole[_T],
) -> roles.ExpressionElementRole[Optional[_T]]: ...


@overload
def Nullable(val: Type[_T]) -> Type[Optional[_T]]: ...


def Nullable(
    val: _TypedColumnClauseArgument[_T],
) -> _TypedColumnClauseArgument[Optional[_T]]:
    """Types a column or ORM class as nullable.

    This can be used in select and other contexts to express that the value of
    a column can be null, for example due to an outer join::

        stmt1 = select(A, Nullable(B)).outerjoin(A.bs)
        stmt2 = select(A.data, Nullable(B.data)).outerjoin(A.bs)

    At runtime this method returns the input unchanged.

    .. versionadded:: 2.0.20
    """
    return val


@overload
def NotNullable(
    val: "SQLCoreOperations[Optional[_T]]",
) -> "SQLCoreOperations[_T]": ...


@overload
def NotNullable(
    val: roles.ExpressionElementRole[Optional[_T]],
) -> roles.ExpressionElementRole[_T]: ...


@overload
def NotNullable(val: Type[Optional[_T]]) -> Type[_T]: ...


@overload
def NotNullable(val: Optional[Type[_T]]) -> Type[_T]: ...


def NotNullable(
    val: Union[_TypedColumnClauseArgument[Optional[_T]], Optional[Type[_T]]],
) -> _TypedColumnClauseArgument[_T]:
    """Types a column or ORM class as not nullable.

    This can be used in select and other contexts to express that the value of
    a column cannot be null, for example due to a where condition on a
    nullable column::

        stmt = select(NotNullable(A.value)).where(A.value.is_not(None))

    At runtime this method returns the input unchanged.

    .. versionadded:: 2.0.20
    """
    return val  # type: ignore

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\testing\__init__.py ===
# testing/__init__.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors


from unittest import mock

from . import config
from .assertions import assert_raises
from .assertions import assert_raises_context_ok
from .assertions import assert_raises_message
from .assertions import assert_raises_message_context_ok
from .assertions import assert_warns
from .assertions import assert_warns_message
from .assertions import AssertsCompiledSQL
from .assertions import AssertsExecutionResults
from .assertions import ComparesIndexes
from .assertions import ComparesTables
from .assertions import emits_warning
from .assertions import emits_warning_on
from .assertions import eq_
from .assertions import eq_ignore_whitespace
from .assertions import eq_regex
from .assertions import expect_deprecated
from .assertions import expect_deprecated_20
from .assertions import expect_raises
from .assertions import expect_raises_message
from .assertions import expect_warnings
from .assertions import in_
from .assertions import int_within_variance
from .assertions import is_
from .assertions import is_false
from .assertions import is_instance_of
from .assertions import is_none
from .assertions import is_not
from .assertions import is_not_
from .assertions import is_not_none
from .assertions import is_true
from .assertions import le_
from .assertions import ne_
from .assertions import not_in
from .assertions import not_in_
from .assertions import startswith_
from .assertions import uses_deprecated
from .config import add_to_marker
from .config import async_test
from .config import combinations
from .config import combinations_list
from .config import db
from .config import fixture
from .config import requirements as requires
from .config import skip_test
from .config import Variation
from .config import variation
from .config import variation_fixture
from .exclusions import _is_excluded
from .exclusions import _server_version
from .exclusions import against as _against
from .exclusions import db_spec
from .exclusions import exclude
from .exclusions import fails
from .exclusions import fails_if
from .exclusions import fails_on
from .exclusions import fails_on_everything_except
from .exclusions import future
from .exclusions import only_if
from .exclusions import only_on
from .exclusions import skip
from .exclusions import skip_if
from .schema import eq_clause_element
from .schema import eq_type_affinity
from .util import adict
from .util import fail
from .util import flag_combinations
from .util import force_drop_names
from .util import lambda_combinations
from .util import metadata_fixture
from .util import provide_metadata
from .util import resolve_lambda
from .util import rowset
from .util import run_as_contextmanager
from .util import skip_if_timeout
from .util import teardown_events
from .warnings import assert_warnings
from .warnings import warn_test_suite


def against(*queries):
    return _against(config._current, *queries)


crashes = skip

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\whisper\whisper_decoder.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.  See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------

import logging
import os
import tempfile
from itertools import chain
from pathlib import Path

import numpy as np
import onnx
import torch
from float16 import convert_float_to_float16
from google.protobuf.internal.containers import RepeatedCompositeFieldContainer
from onnx import ModelProto, ValueInfoProto
from onnx_model import OnnxModel
from past_helper import PastKeyValuesHelper
from transformers import WhisperConfig
from whisper_inputs import (
    convert_inputs_for_ort,
    get_model_dynamic_axes,
    get_sample_decoder_inputs,
    group_past_key_values,
)

from onnxruntime import InferenceSession

logger = logging.getLogger(__name__)


class WhisperDecoder(torch.nn.Module):
    """A Whisper decoder with optional past key values"""

    def __init__(self, config: WhisperConfig, model: torch.nn.Module, model_impl: str, no_beam_search_op: bool = False):
        super().__init__()
        self.config = config
        self.device = model.device
        self.model_impl = model_impl
        self.no_beam_search_op = no_beam_search_op

        self.decoder = None if model_impl == "openai" else model.model.decoder
        self.proj_out = None if model_impl == "openai" else model.proj_out
        self.model = model if model_impl == "openai" else None

        self.max_source_positions = self.config.max_source_positions
        self.num_heads = self.config.decoder_attention_heads
        self.head_size = self.config.d_model // self.num_heads

    def hf_forward(
        self,
        decoder_input_ids: torch.Tensor,
        encoder_hidden_states: torch.Tensor | None = None,
        past_key_values: list[tuple[torch.Tensor]] | None = None,
    ):
        outputs = self.decoder(
            encoder_hidden_states=encoder_hidden_states,
            input_ids=decoder_input_ids,
            past_key_values=past_key_values,
            use_cache=True,
        )
        logits = self.proj_out(outputs.last_hidden_state)
        present_key_values = outputs.past_key_values

        if past_key_values is None:
            # Return present_self_* and present_cross_* for decoder-init
            return logits, present_key_values

        # Before: (past_key_self_0, past_value_self_0, past_key_cross_0, past_value_cross_0),
        #         (past_key_self_1, past_value_self_1, past_key_cross_1, past_value_cross_1),
        # After:  (past_key_self_0, past_value_self_0, past_key_self_1, past_value_self_1), ...,
        #         (past_key_cross_0, past_value_cross_0, past_key_cross_1, past_value_cross_1), ...
        present_self, present_cross = PastKeyValuesHelper.group_by_self_and_cross(present_key_values)

        # Return present_self_* for decoder-with-past since past_cross_* and present_cross_* are identical
        return logits, present_self

    def oai_forward(
        self,
        decoder_input_ids: torch.Tensor,
        encoder_hidden_states: torch.Tensor | None = None,
        past_key_values: list[tuple[torch.Tensor]] | None = None,
    ):
        past_kv_cache = {}
        if past_key_values is not None:
            # Convert past KV caches (BxNxSxH --> BxSxNxH --> BxSxD) for OpenAI's forward pass
            self_attn_kv_caches, cross_attn_kv_caches = group_past_key_values(past_key_values)
            self_attn_kv_caches = [past_kv.transpose(1, 2) for past_kv in self_attn_kv_caches]
            self_attn_kv_caches = [past_kv.reshape(past_kv.shape[:2] + (-1,)) for past_kv in self_attn_kv_caches]
            cross_attn_kv_caches = [past_kv.transpose(1, 2) for past_kv in cross_attn_kv_caches]
            cross_attn_kv_caches = [past_kv.reshape(past_kv.shape[:2] + (-1,)) for past_kv in cross_attn_kv_caches]

            for idx, block in enumerate(self.model.decoder.blocks):
                past_kv_cache[block.attn.key] = self_attn_kv_caches[2 * idx]
                past_kv_cache[block.attn.value] = self_attn_kv_caches[2 * idx + 1]
                past_kv_cache[block.cross_attn.key] = cross_attn_kv_caches[2 * idx]
                past_kv_cache[block.cross_attn.value] = cross_attn_kv_caches[2 * idx + 1]

        # Install OpenAI's hooks on the forward pass of each nn.Linear for key and value
        # since the hooks will capture the output of the key and value MatMuls, which
        # represent the current keys and values.
        #
        # For OpenAI's forward pass, the hook function will also perform the concat
        # operation (past_kv + curr_kv --> pres_kv) if needed. However, the ONNX model
        # will not contain this concat operation because the present KV caches aren't
        # returned by OpenAI's forward pass.
        kv_cache, hooks = self.model.install_kv_cache_hooks()

        # Run forward pass
        # NOTE: There is a bug with openai-whisper==20240930 with the introduction of SDPA.
        # In the Whisper codebase, the following line
        #
        # is_causal = mask is not None and n_ctx > 1
        #
        # has been added where `mask` is a torch tensor. The right-hand side evaluates to `tensor(True/False)`
        # but `is_causal` only accepts the boolean value. The fix is to apply `.item()` after the right-hand
        # side has been evaluated. In other words, the line should be
        #
        # is_causal = (mask is not None and n_ctx > 1).item()
        #
        # instead.
        logits = self.model.decoder(x=decoder_input_ids, xa=encoder_hidden_states, kv_cache=past_kv_cache)

        # Re-do concat operation on self attention KV caches for ONNX export (if past self attention KV caches exist)
        if past_key_values is not None:
            for block in self.model.decoder.blocks:
                kv_cache[block.attn.key] = torch.cat(
                    [past_kv_cache[block.attn.key], kv_cache[block.attn.key]], dim=1
                ).detach()
                kv_cache[block.attn.value] = torch.cat(
                    [past_kv_cache[block.attn.value], kv_cache[block.attn.value]], dim=1
                ).detach()

        present_self, present_cross = [], []
        for block in self.model.decoder.blocks:
            # Group self and cross values
            present_self.append(kv_cache[block.attn.key])
            present_self.append(kv_cache[block.attn.value])
            if past_key_values is None:
                # Return present_self_* and present_cross_* for decoder-init
                present_cross.append(kv_cache[block.cross_attn.key])
                present_cross.append(kv_cache[block.cross_attn.value])

        # Convert present KV caches (BxSxD --> BxSxNxH --> BxNxSxH) after OpenAI's forward pass
        present_self = [
            present_kv.reshape(present_kv.shape[:2] + (-1, self.head_size)).transpose(1, 2)
            for present_kv in present_self
        ]
        present_cross = [
            present_kv.reshape(present_kv.shape[:2] + (-1, self.head_size)).transpose(1, 2)
            for present_kv in present_cross
        ]

        # Remove OpenAI's hooks since they can persist after this function completes
        for hook in hooks:
            hook.remove()

        if past_key_values is None:
            # Return present_self_* and present_cross_* for decoder-init
            present_key_values = PastKeyValuesHelper.group_by_layer(
                present_self + present_cross, len(present_self) // 2
            )
            return logits, present_key_values

        # Return present_self_* for decoder-with-past since past_cross_* and present_cross_* are identical
        return logits, present_self

    def forward(
        self,
        decoder_input_ids: torch.Tensor,
        encoder_hidden_states: torch.Tensor | None = None,
        past_key_values: list[tuple[torch.Tensor]] | None = None,
    ):
        if self.model_impl == "openai":
            return self.oai_forward(decoder_input_ids, encoder_hidden_states, past_key_values)
        return self.hf_forward(decoder_input_ids, encoder_hidden_states, past_key_values)

    def input_names(self):
        if self.first_pass:
            input_names = ["input_ids", "encoder_hidden_states"]
        else:
            input_names = [
                "input_ids",
                "encoder_hidden_states",
                *list(
                    chain.from_iterable(
                        (f"past_key_self_{i}", f"past_value_self_{i}", f"past_key_cross_{i}", f"past_value_cross_{i}")
                        for i in range(self.config.num_hidden_layers)
                    )
                ),
            ]
        return input_names

    def output_names(self):
        if self.first_pass:
            output_names = [
                "logits",
                *list(
                    chain.from_iterable(
                        (
                            f"present_key_self_{i}",
                            f"present_value_self_{i}",
                            f"present_key_cross_{i}",
                            f"present_value_cross_{i}",
                        )
                        for i in range(self.config.num_hidden_layers)
                    )
                ),
            ]
        else:
            output_names = [
                "logits",
                *list(
                    chain.from_iterable(
                        (f"present_key_self_{i}", f"present_value_self_{i}")
                        for i in range(self.config.num_hidden_layers)
                    )
                ),
            ]
        return output_names

    def dynamic_axes(self, input_names, output_names):
        dynamic_axes = get_model_dynamic_axes(self.config, input_names, output_names)
        if "input_ids" in dynamic_axes and not self.no_beam_search_op:
            # Set dynamic axes for `input_ids` when using beam search op to {0: "batch_size"} only
            del dynamic_axes["input_ids"][1]
        return dynamic_axes

    def inputs(self, use_fp16_inputs: bool, use_int32_inputs: bool, return_dict: bool = False):
        inputs = get_sample_decoder_inputs(
            self.config,
            self.device,
            batch_size=2,
            past_sequence_length=(0 if self.first_pass else 6),
            sequence_length=(6 if self.first_pass else 1),
            use_fp16=use_fp16_inputs,
            use_int32=use_int32_inputs,
        )
        if return_dict:
            if self.first_pass:
                del inputs["past_key_values"]
            return inputs

        if self.first_pass:
            return (
                inputs["decoder_input_ids"],
                inputs["encoder_hidden_states"],
            )
        return (
            inputs["decoder_input_ids"],
            inputs["encoder_hidden_states"],
            inputs["past_key_values"],
        )

    def fix_key_value_cache_dims(self, io: ValueInfoProto, is_cross: bool = False, is_output: bool = False):
        # Shape should be (batch_size, num_heads, sequence_length, head_size) for self attention KV caches
        # and (batch_size, num_heads, num_frames // 2, head_size) for cross attention KV caches
        num_heads = io.type.tensor_type.shape.dim[1]
        if "_dim_" in num_heads.dim_param:
            num_heads.Clear()
            num_heads.dim_value = self.num_heads
        sequence_length = io.type.tensor_type.shape.dim[2]
        if "_dim_" in sequence_length.dim_param:
            sequence_length.Clear()
            if is_cross:
                sequence_length.dim_value = self.max_source_positions
            else:
                sequence_length.dim_param = "total_sequence_length" if is_output else "past_sequence_length"
        head_size = io.type.tensor_type.shape.dim[3]
        if "_dim_" in head_size.dim_param:
            head_size.Clear()
            head_size.dim_value = self.head_size
        return io

    def fix_io(self, io_list: RepeatedCompositeFieldContainer, is_output: bool = False):
        # Fix order of inputs/outputs and each dim_value of input/output
        reordered_io = []
        self_attn_kv_caches = []
        cross_attn_kv_caches = []

        for io in io_list:
            if "past" not in io.name and "present" not in io.name:
                reordered_io.append(io)
            elif "self" in io.name:
                # Self attention KV caches
                new_io = self.fix_key_value_cache_dims(io, is_cross=False, is_output=is_output)
                if self.no_beam_search_op:
                    reordered_io.append(new_io)
                else:
                    self_attn_kv_caches.append(new_io)
            else:
                # Cross attention KV caches
                new_io = self.fix_key_value_cache_dims(io, is_cross=True, is_output=is_output)
                if self.no_beam_search_op:
                    reordered_io.append(new_io)
                else:
                    cross_attn_kv_caches.append(new_io)

        if not self.no_beam_search_op:
            reordered_io += self_attn_kv_caches + cross_attn_kv_caches
        return reordered_io

    def fix_inputs_and_outputs(self, model: ModelProto):
        # ONNX exporter might mark dimensions like 'Transposepresent_value_self_1_dim_2' in shape inference.
        # We now change the dim_values to the correct one.
        reordered_inputs = self.fix_io(model.graph.input, is_output=False)
        while len(model.graph.input) > 0:
            model.graph.input.pop()
        model.graph.input.extend(reordered_inputs)

        reordered_outputs = self.fix_io(model.graph.output, is_output=True)
        while len(model.graph.output) > 0:
            model.graph.output.pop()
        model.graph.output.extend(reordered_outputs)
        return model

    def fix_layernorm_weights(self, model: ModelProto, use_fp16_inputs: bool):
        if self.model_impl == "openai" and use_fp16_inputs:
            # Cast ONNX model to float16 to ensure LayerNorm weights are converted from
            # float32 to float16 since exported model already has float16 weights everywhere
            # except for LayerNorm ops. This happens because OpenAI always upcasts to float32
            # when computing LayerNorm.
            #
            # Reference:
            # https://github.com/openai/whisper/blob/90db0de1896c23cbfaf0c58bc2d30665f709f170/whisper/model.py#L41
            model = convert_float_to_float16(model)
        return model

    def export_onnx(
        self,
        onnx_model_path: str,
        provider: str,
        verbose: bool = True,
        use_external_data_format: bool = False,
        use_fp16_inputs: bool = False,
        use_int32_inputs: bool = True,
        use_encoder_hidden_states: bool = False,
        use_kv_cache_inputs: bool = True,
    ):
        """Export decoder to ONNX

        Args:
            onnx_model_path (str): path to save ONNX model
            provider (str): provider to use for verifying parity on ONNX model
            verbose (bool, optional): print verbose information. Defaults to True.
            use_external_data_format (bool, optional): use external data format or not. Defaults to False.
            use_fp16_inputs (bool, optional): use float16 inputs for the KV caches. Defaults to False.
            use_int32_inputs (bool, optional): use int32 inputs for the decoder_input_ids. Defaults to True.
            use_encoder_hidden_states (bool, optional): use encoder_hidden_states as model input for decoder-init/decoder-without-past models. Defaults to False.
            use_kv_cache_inputs (bool, optional): use KV caches as model inputs for decoder-with-past models. Defaults to True.
        """
        # Shape of decoder's tensors:
        # Required Inputs:
        #    decoder_input_ids: (batch_size, sequence_length)
        # Optional Inputs:
        #    encoder_hidden_states (comes from encoder's outputs): (batch_size, num_frames // 2, hidden_size)
        #    past_{key/value}_self_* (past self attention KV caches): (batch_size, num_heads, past_sequence_length, head_size)
        #    past_{key/value}_cross_* (past cross attention KV caches): (batch_size, num_heads, num_frames // 2, head_size)
        # Outputs:
        #    logits: (batch_size, sequence_length, vocab_size)
        #    present_{key/value}_self_* (present self attention KV caches): (batch_size, num_heads, past_sequence_length + sequence_length, head_size)
        #    present_{key/value}_cross_* (present cross attention KV caches): (batch_size, num_heads, num_frames // 2, head_size)

        # For the first pass through the decoder (i.e. decoder-init/decoder-without-past)
        self.first_pass = use_encoder_hidden_states and not use_kv_cache_inputs

        # For subsequent passes through the decoder (i.e. decoder-with-past)
        self.later_pass = not use_encoder_hidden_states and use_kv_cache_inputs

        assert self.first_pass or self.later_pass, (
            "Only one of `use_encoder_hidden_states` and `use_kv_cache_inputs` can be true at once."
        )

        inputs = self.inputs(use_fp16_inputs=use_fp16_inputs, use_int32_inputs=use_int32_inputs)
        input_names = self.input_names()
        output_names = self.output_names()
        dynamic_axes = self.dynamic_axes(input_names, output_names)

        Path(onnx_model_path).parent.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory() as tmp_dir_name:
            temp_onnx_model_path = os.path.join(tmp_dir_name, "decoder.onnx")
            Path(temp_onnx_model_path).parent.mkdir(parents=True, exist_ok=True)
            out_path = temp_onnx_model_path if use_external_data_format else onnx_model_path

            torch.onnx.export(
                self,
                args=inputs,
                f=out_path,
                export_params=True,
                input_names=input_names,
                output_names=output_names,
                dynamic_axes=dynamic_axes,
                opset_version=17,
                do_constant_folding=True,
                verbose=verbose,
            )

            model = onnx.load_model(out_path, load_external_data=use_external_data_format)
            model = self.fix_inputs_and_outputs(model)
            model = self.fix_layernorm_weights(model, use_fp16_inputs)
            OnnxModel.save(
                model,
                onnx_model_path,
                save_as_external_data=use_external_data_format,
                all_tensors_to_one_file=True,
            )

        self.verify_onnx(onnx_model_path, provider, use_fp16_inputs, use_int32_inputs)

    def verify_onnx(
        self,
        onnx_model_path: str,
        provider: str,
        use_fp16_inputs: bool,
        use_int32_inputs: bool,
    ):
        """Verify ONNX model outputs and PyTorch model outputs match

        Args:
            onnx_model_path (str): path to save ONNX model
            provider (str): execution provider for ONNX model
            use_fp16_inputs (bool, optional): use float16 inputs for the KV caches
            use_int32_inputs (bool, optional): use int32 inputs for the decoder_input_ids
        """
        # Shape of decoder's tensors:
        # Required Inputs:
        #    decoder_input_ids: (batch_size, sequence_length)
        # Optional Inputs:
        #    encoder_hidden_states (comes from encoder's outputs): (batch_size, num_frames // 2, hidden_size)
        #    past_{key/value}_self_* (past self attention KV caches): (batch_size, num_heads, past_sequence_length, head_size)
        #    past_{key/value}_cross_* (past cross attention KV caches): (batch_size, num_heads, num_frames // 2, head_size)
        # Outputs:
        #    logits: (batch_size, sequence_length, vocab_size)
        #    present_{key/value}_self_* (present self attention KV caches): (batch_size, num_heads, past_sequence_length + sequence_length, head_size)
        #    present_{key/value}_cross_* (present cross attention KV caches): (batch_size, num_heads, num_frames // 2, head_size)

        # Run PyTorch model
        inputs = self.inputs(use_fp16_inputs=use_fp16_inputs, use_int32_inputs=use_int32_inputs, return_dict=True)
        pt_outputs = []
        if self.first_pass:
            out = self.forward(**inputs)
            pt_outputs.append(out[0].detach().cpu().numpy())
            for present_key_value_layer in out[1]:
                for present_key_value in present_key_value_layer:
                    pt_outputs.append(present_key_value.detach().cpu().numpy())
        else:
            out = self.forward(**inputs)
            pt_outputs.append(out[0].detach().cpu().numpy())
            for present_self_key_value in out[1]:
                pt_outputs.append(present_self_key_value.detach().cpu().numpy())

        # Run ONNX model
        sess = InferenceSession(onnx_model_path, providers=[provider])
        ort_outputs = sess.run(None, convert_inputs_for_ort(inputs, sess))

        # Calculate output difference
        try:
            for i, output_name in enumerate(self.output_names()):
                diff = np.abs(pt_outputs[i] - ort_outputs[i])
                logger.warning(f"Comparing {output_name}...")
                logger.warning(f"Max diff: {np.max(diff)}")
        except:  # noqa: E722
            pass

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\interchange\dataframe_protocol.py ===
"""
A verbatim copy (vendored) of the spec from https://github.com/data-apis/dataframe-api
"""

from __future__ import annotations

from abc import (
    ABC,
    abstractmethod,
)
import enum
from typing import (
    TYPE_CHECKING,
    Any,
    TypedDict,
)

if TYPE_CHECKING:
    from collections.abc import (
        Iterable,
        Sequence,
    )


class DlpackDeviceType(enum.IntEnum):
    """Integer enum for device type codes matching DLPack."""

    CPU = 1
    CUDA = 2
    CPU_PINNED = 3
    OPENCL = 4
    VULKAN = 7
    METAL = 8
    VPI = 9
    ROCM = 10


class DtypeKind(enum.IntEnum):
    """
    Integer enum for data types.

    Attributes
    ----------
    INT : int
        Matches to signed integer data type.
    UINT : int
        Matches to unsigned integer data type.
    FLOAT : int
        Matches to floating point data type.
    BOOL : int
        Matches to boolean data type.
    STRING : int
        Matches to string data type (UTF-8 encoded).
    DATETIME : int
        Matches to datetime data type.
    CATEGORICAL : int
        Matches to categorical data type.
    """

    INT = 0
    UINT = 1
    FLOAT = 2
    BOOL = 20
    STRING = 21  # UTF-8
    DATETIME = 22
    CATEGORICAL = 23


class ColumnNullType(enum.IntEnum):
    """
    Integer enum for null type representation.

    Attributes
    ----------
    NON_NULLABLE : int
        Non-nullable column.
    USE_NAN : int
        Use explicit float NaN value.
    USE_SENTINEL : int
        Sentinel value besides NaN/NaT.
    USE_BITMASK : int
        The bit is set/unset representing a null on a certain position.
    USE_BYTEMASK : int
        The byte is set/unset representing a null on a certain position.
    """

    NON_NULLABLE = 0
    USE_NAN = 1
    USE_SENTINEL = 2
    USE_BITMASK = 3
    USE_BYTEMASK = 4


class ColumnBuffers(TypedDict):
    # first element is a buffer containing the column data;
    # second element is the data buffer's associated dtype
    data: tuple[Buffer, Any]

    # first element is a buffer containing mask values indicating missing data;
    # second element is the mask value buffer's associated dtype.
    # None if the null representation is not a bit or byte mask
    validity: tuple[Buffer, Any] | None

    # first element is a buffer containing the offset values for
    # variable-size binary data (e.g., variable-length strings);
    # second element is the offsets buffer's associated dtype.
    # None if the data buffer does not have an associated offsets buffer
    offsets: tuple[Buffer, Any] | None


class CategoricalDescription(TypedDict):
    # whether the ordering of dictionary indices is semantically meaningful
    is_ordered: bool
    # whether a dictionary-style mapping of categorical values to other objects exists
    is_dictionary: bool
    # Python-level only (e.g. ``{int: str}``).
    # None if not a dictionary-style categorical.
    categories: Column | None


class Buffer(ABC):
    """
    Data in the buffer is guaranteed to be contiguous in memory.

    Note that there is no dtype attribute present, a buffer can be thought of
    as simply a block of memory. However, if the column that the buffer is
    attached to has a dtype that's supported by DLPack and ``__dlpack__`` is
    implemented, then that dtype information will be contained in the return
    value from ``__dlpack__``.

    This distinction is useful to support both data exchange via DLPack on a
    buffer and (b) dtypes like variable-length strings which do not have a
    fixed number of bytes per element.
    """

    @property
    @abstractmethod
    def bufsize(self) -> int:
        """
        Buffer size in bytes.
        """

    @property
    @abstractmethod
    def ptr(self) -> int:
        """
        Pointer to start of the buffer as an integer.
        """

    @abstractmethod
    def __dlpack__(self):
        """
        Produce DLPack capsule (see array API standard).

        Raises:

            - TypeError : if the buffer contains unsupported dtypes.
            - NotImplementedError : if DLPack support is not implemented

        Useful to have to connect to array libraries. Support optional because
        it's not completely trivial to implement for a Python-only library.
        """
        raise NotImplementedError("__dlpack__")

    @abstractmethod
    def __dlpack_device__(self) -> tuple[DlpackDeviceType, int | None]:
        """
        Device type and device ID for where the data in the buffer resides.
        Uses device type codes matching DLPack.
        Note: must be implemented even if ``__dlpack__`` is not.
        """


class Column(ABC):
    """
    A column object, with only the methods and properties required by the
    interchange protocol defined.

    A column can contain one or more chunks. Each chunk can contain up to three
    buffers - a data buffer, a mask buffer (depending on null representation),
    and an offsets buffer (if variable-size binary; e.g., variable-length
    strings).

    TBD: Arrow has a separate "null" dtype, and has no separate mask concept.
         Instead, it seems to use "children" for both columns with a bit mask,
         and for nested dtypes. Unclear whether this is elegant or confusing.
         This design requires checking the null representation explicitly.

         The Arrow design requires checking:
         1. the ARROW_FLAG_NULLABLE (for sentinel values)
         2. if a column has two children, combined with one of those children
            having a null dtype.

         Making the mask concept explicit seems useful. One null dtype would
         not be enough to cover both bit and byte masks, so that would mean
         even more checking if we did it the Arrow way.

    TBD: there's also the "chunk" concept here, which is implicit in Arrow as
         multiple buffers per array (= column here). Semantically it may make
         sense to have both: chunks were meant for example for lazy evaluation
         of data which doesn't fit in memory, while multiple buffers per column
         could also come from doing a selection operation on a single
         contiguous buffer.

         Given these concepts, one would expect chunks to be all of the same
         size (say a 10,000 row dataframe could have 10 chunks of 1,000 rows),
         while multiple buffers could have data-dependent lengths. Not an issue
         in pandas if one column is backed by a single NumPy array, but in
         Arrow it seems possible.
         Are multiple chunks *and* multiple buffers per column necessary for
         the purposes of this interchange protocol, or must producers either
         reuse the chunk concept for this or copy the data?

    Note: this Column object can only be produced by ``__dataframe__``, so
          doesn't need its own version or ``__column__`` protocol.
    """

    @abstractmethod
    def size(self) -> int:
        """
        Size of the column, in elements.

        Corresponds to DataFrame.num_rows() if column is a single chunk;
        equal to size of this current chunk otherwise.
        """

    @property
    @abstractmethod
    def offset(self) -> int:
        """
        Offset of first element.

        May be > 0 if using chunks; for example for a column with N chunks of
        equal size M (only the last chunk may be shorter),
        ``offset = n * M``, ``n = 0 .. N-1``.
        """

    @property
    @abstractmethod
    def dtype(self) -> tuple[DtypeKind, int, str, str]:
        """
        Dtype description as a tuple ``(kind, bit-width, format string, endianness)``.

        Bit-width : the number of bits as an integer
        Format string : data type description format string in Apache Arrow C
                        Data Interface format.
        Endianness : current only native endianness (``=``) is supported

        Notes:
            - Kind specifiers are aligned with DLPack where possible (hence the
              jump to 20, leave enough room for future extension)
            - Masks must be specified as boolean with either bit width 1 (for bit
              masks) or 8 (for byte masks).
            - Dtype width in bits was preferred over bytes
            - Endianness isn't too useful, but included now in case in the future
              we need to support non-native endianness
            - Went with Apache Arrow format strings over NumPy format strings
              because they're more complete from a dataframe perspective
            - Format strings are mostly useful for datetime specification, and
              for categoricals.
            - For categoricals, the format string describes the type of the
              categorical in the data buffer. In case of a separate encoding of
              the categorical (e.g. an integer to string mapping), this can
              be derived from ``self.describe_categorical``.
            - Data types not included: complex, Arrow-style null, binary, decimal,
              and nested (list, struct, map, union) dtypes.
        """

    @property
    @abstractmethod
    def describe_categorical(self) -> CategoricalDescription:
        """
        If the dtype is categorical, there are two options:
        - There are only values in the data buffer.
        - There is a separate non-categorical Column encoding for categorical values.

        Raises TypeError if the dtype is not categorical

        Returns the dictionary with description on how to interpret the data buffer:
            - "is_ordered" : bool, whether the ordering of dictionary indices is
                             semantically meaningful.
            - "is_dictionary" : bool, whether a mapping of
                                categorical values to other objects exists
            - "categories" : Column representing the (implicit) mapping of indices to
                             category values (e.g. an array of cat1, cat2, ...).
                             None if not a dictionary-style categorical.

        TBD: are there any other in-memory representations that are needed?
        """

    @property
    @abstractmethod
    def describe_null(self) -> tuple[ColumnNullType, Any]:
        """
        Return the missing value (or "null") representation the column dtype
        uses, as a tuple ``(kind, value)``.

        Value : if kind is "sentinel value", the actual value. If kind is a bit
        mask or a byte mask, the value (0 or 1) indicating a missing value. None
        otherwise.
        """

    @property
    @abstractmethod
    def null_count(self) -> int | None:
        """
        Number of null elements, if known.

        Note: Arrow uses -1 to indicate "unknown", but None seems cleaner.
        """

    @property
    @abstractmethod
    def metadata(self) -> dict[str, Any]:
        """
        The metadata for the column. See `DataFrame.metadata` for more details.
        """

    @abstractmethod
    def num_chunks(self) -> int:
        """
        Return the number of chunks the column consists of.
        """

    @abstractmethod
    def get_chunks(self, n_chunks: int | None = None) -> Iterable[Column]:
        """
        Return an iterator yielding the chunks.

        See `DataFrame.get_chunks` for details on ``n_chunks``.
        """

    @abstractmethod
    def get_buffers(self) -> ColumnBuffers:
        """
        Return a dictionary containing the underlying buffers.

        The returned dictionary has the following contents:

            - "data": a two-element tuple whose first element is a buffer
                      containing the data and whose second element is the data
                      buffer's associated dtype.
            - "validity": a two-element tuple whose first element is a buffer
                          containing mask values indicating missing data and
                          whose second element is the mask value buffer's
                          associated dtype. None if the null representation is
                          not a bit or byte mask.
            - "offsets": a two-element tuple whose first element is a buffer
                         containing the offset values for variable-size binary
                         data (e.g., variable-length strings) and whose second
                         element is the offsets buffer's associated dtype. None
                         if the data buffer does not have an associated offsets
                         buffer.
        """


#    def get_children(self) -> Iterable[Column]:
#        """
#        Children columns underneath the column, each object in this iterator
#        must adhere to the column specification.
#        """
#        pass


class DataFrame(ABC):
    """
    A data frame class, with only the methods required by the interchange
    protocol defined.

    A "data frame" represents an ordered collection of named columns.
    A column's "name" must be a unique string.
    Columns may be accessed by name or by position.

    This could be a public data frame class, or an object with the methods and
    attributes defined on this DataFrame class could be returned from the
    ``__dataframe__`` method of a public data frame class in a library adhering
    to the dataframe interchange protocol specification.
    """

    version = 0  # version of the protocol

    @abstractmethod
    def __dataframe__(self, nan_as_null: bool = False, allow_copy: bool = True):
        """Construct a new interchange object, potentially changing the parameters."""

    @property
    @abstractmethod
    def metadata(self) -> dict[str, Any]:
        """
        The metadata for the data frame, as a dictionary with string keys. The
        contents of `metadata` may be anything, they are meant for a library
        to store information that it needs to, e.g., roundtrip losslessly or
        for two implementations to share data that is not (yet) part of the
        interchange protocol specification. For avoiding collisions with other
        entries, please add name the keys with the name of the library
        followed by a period and the desired name, e.g, ``pandas.indexcol``.
        """

    @abstractmethod
    def num_columns(self) -> int:
        """
        Return the number of columns in the DataFrame.
        """

    @abstractmethod
    def num_rows(self) -> int | None:
        # TODO: not happy with Optional, but need to flag it may be expensive
        #       why include it if it may be None - what do we expect consumers
        #       to do here?
        """
        Return the number of rows in the DataFrame, if available.
        """

    @abstractmethod
    def num_chunks(self) -> int:
        """
        Return the number of chunks the DataFrame consists of.
        """

    @abstractmethod
    def column_names(self) -> Iterable[str]:
        """
        Return an iterator yielding the column names.
        """

    @abstractmethod
    def get_column(self, i: int) -> Column:
        """
        Return the column at the indicated position.
        """

    @abstractmethod
    def get_column_by_name(self, name: str) -> Column:
        """
        Return the column whose name is the indicated name.
        """

    @abstractmethod
    def get_columns(self) -> Iterable[Column]:
        """
        Return an iterator yielding the columns.
        """

    @abstractmethod
    def select_columns(self, indices: Sequence[int]) -> DataFrame:
        """
        Create a new DataFrame by selecting a subset of columns by index.
        """

    @abstractmethod
    def select_columns_by_name(self, names: Sequence[str]) -> DataFrame:
        """
        Create a new DataFrame by selecting a subset of columns by name.
        """

    @abstractmethod
    def get_chunks(self, n_chunks: int | None = None) -> Iterable[DataFrame]:
        """
        Return an iterator yielding the chunks.

        By default (None), yields the chunks that the data is stored as by the
        producer. If given, ``n_chunks`` must be a multiple of
        ``self.num_chunks()``, meaning the producer must subdivide each chunk
        before yielding it.
        """

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\expressions\hadamard.py ===
from collections import Counter

from sympy.core import Mul, sympify
from sympy.core.add import Add
from sympy.core.expr import ExprBuilder
from sympy.core.sorting import default_sort_key
from sympy.functions.elementary.exponential import log
from sympy.matrices.expressions.matexpr import MatrixExpr
from sympy.matrices.expressions._shape import validate_matadd_integer as validate
from sympy.matrices.expressions.special import ZeroMatrix, OneMatrix
from sympy.strategies import (
    unpack, flatten, condition, exhaust, rm_id, sort
)
from sympy.utilities.exceptions import sympy_deprecation_warning


def hadamard_product(*matrices):
    """
    Return the elementwise (aka Hadamard) product of matrices.

    Examples
    ========

    >>> from sympy import hadamard_product, MatrixSymbol
    >>> A = MatrixSymbol('A', 2, 3)
    >>> B = MatrixSymbol('B', 2, 3)
    >>> hadamard_product(A)
    A
    >>> hadamard_product(A, B)
    HadamardProduct(A, B)
    >>> hadamard_product(A, B)[0, 1]
    A[0, 1]*B[0, 1]
    """
    if not matrices:
        raise TypeError("Empty Hadamard product is undefined")
    if len(matrices) == 1:
        return matrices[0]
    return HadamardProduct(*matrices).doit()


class HadamardProduct(MatrixExpr):
    """
    Elementwise product of matrix expressions

    Examples
    ========

    Hadamard product for matrix symbols:

    >>> from sympy import hadamard_product, HadamardProduct, MatrixSymbol
    >>> A = MatrixSymbol('A', 5, 5)
    >>> B = MatrixSymbol('B', 5, 5)
    >>> isinstance(hadamard_product(A, B), HadamardProduct)
    True

    Notes
    =====

    This is a symbolic object that simply stores its argument without
    evaluating it. To actually compute the product, use the function
    ``hadamard_product()`` or ``HadamardProduct.doit``
    """
    is_HadamardProduct = True

    def __new__(cls, *args, evaluate=False, check=None):
        args = list(map(sympify, args))
        if len(args) == 0:
            # We currently don't have a way to support one-matrices of generic dimensions:
            raise ValueError("HadamardProduct needs at least one argument")

        if not all(isinstance(arg, MatrixExpr) for arg in args):
            raise TypeError("Mix of Matrix and Scalar symbols")

        if check is not None:
            sympy_deprecation_warning(
                "Passing check to HadamardProduct is deprecated and the check argument will be removed in a future version.",
                deprecated_since_version="1.11",
                active_deprecations_target='remove-check-argument-from-matrix-operations')

        if check is not False:
            validate(*args)

        obj = super().__new__(cls, *args)
        if evaluate:
            obj = obj.doit(deep=False)
        return obj

    @property
    def shape(self):
        return self.args[0].shape

    def _entry(self, i, j, **kwargs):
        return Mul(*[arg._entry(i, j, **kwargs) for arg in self.args])

    def _eval_transpose(self):
        from sympy.matrices.expressions.transpose import transpose
        return HadamardProduct(*list(map(transpose, self.args)))

    def doit(self, **hints):
        expr = self.func(*(i.doit(**hints) for i in self.args))
        # Check for explicit matrices:
        from sympy.matrices.matrixbase import MatrixBase
        from sympy.matrices.immutable import ImmutableMatrix

        explicit = [i for i in expr.args if isinstance(i, MatrixBase)]
        if explicit:
            remainder = [i for i in expr.args if i not in explicit]
            expl_mat = ImmutableMatrix([
                Mul.fromiter(i) for i in zip(*explicit)
            ]).reshape(*self.shape)
            expr = HadamardProduct(*([expl_mat] + remainder))

        return canonicalize(expr)

    def _eval_derivative(self, x):
        terms = []
        args = list(self.args)
        for i in range(len(args)):
            factors = args[:i] + [args[i].diff(x)] + args[i+1:]
            terms.append(hadamard_product(*factors))
        return Add.fromiter(terms)

    def _eval_derivative_matrix_lines(self, x):
        from sympy.tensor.array.expressions.array_expressions import ArrayDiagonal
        from sympy.tensor.array.expressions.array_expressions import ArrayTensorProduct
        from sympy.matrices.expressions.matexpr import _make_matrix

        with_x_ind = [i for i, arg in enumerate(self.args) if arg.has(x)]
        lines = []
        for ind in with_x_ind:
            left_args = self.args[:ind]
            right_args = self.args[ind+1:]

            d = self.args[ind]._eval_derivative_matrix_lines(x)
            hadam = hadamard_product(*(right_args + left_args))
            diagonal = [(0, 2), (3, 4)]
            diagonal = [e for j, e in enumerate(diagonal) if self.shape[j] != 1]
            for i in d:
                l1 = i._lines[i._first_line_index]
                l2 = i._lines[i._second_line_index]
                subexpr = ExprBuilder(
                    ArrayDiagonal,
                    [
                        ExprBuilder(
                            ArrayTensorProduct,
                            [
                                ExprBuilder(_make_matrix, [l1]),
                                hadam,
                                ExprBuilder(_make_matrix, [l2]),
                            ]
                        ),
                    *diagonal],

                )
                i._first_pointer_parent = subexpr.args[0].args[0].args
                i._first_pointer_index = 0
                i._second_pointer_parent = subexpr.args[0].args[2].args
                i._second_pointer_index = 0
                i._lines = [subexpr]
                lines.append(i)

        return lines


# TODO Implement algorithm for rewriting Hadamard product as diagonal matrix
# if matmul identy matrix is multiplied.
def canonicalize(x):
    """Canonicalize the Hadamard product ``x`` with mathematical properties.

    Examples
    ========

    >>> from sympy import MatrixSymbol, HadamardProduct
    >>> from sympy import OneMatrix, ZeroMatrix
    >>> from sympy.matrices.expressions.hadamard import canonicalize
    >>> from sympy import init_printing
    >>> init_printing(use_unicode=False)

    >>> A = MatrixSymbol('A', 2, 2)
    >>> B = MatrixSymbol('B', 2, 2)
    >>> C = MatrixSymbol('C', 2, 2)

    Hadamard product associativity:

    >>> X = HadamardProduct(A, HadamardProduct(B, C))
    >>> X
    A.*(B.*C)
    >>> canonicalize(X)
    A.*B.*C

    Hadamard product commutativity:

    >>> X = HadamardProduct(A, B)
    >>> Y = HadamardProduct(B, A)
    >>> X
    A.*B
    >>> Y
    B.*A
    >>> canonicalize(X)
    A.*B
    >>> canonicalize(Y)
    A.*B

    Hadamard product identity:

    >>> X = HadamardProduct(A, OneMatrix(2, 2))
    >>> X
    A.*1
    >>> canonicalize(X)
    A

    Absorbing element of Hadamard product:

    >>> X = HadamardProduct(A, ZeroMatrix(2, 2))
    >>> X
    A.*0
    >>> canonicalize(X)
    0

    Rewriting to Hadamard Power

    >>> X = HadamardProduct(A, A, A)
    >>> X
    A.*A.*A
    >>> canonicalize(X)
     .3
    A

    Notes
    =====

    As the Hadamard product is associative, nested products can be flattened.

    The Hadamard product is commutative so that factors can be sorted for
    canonical form.

    A matrix of only ones is an identity for Hadamard product,
    so every matrices of only ones can be removed.

    Any zero matrix will make the whole product a zero matrix.

    Duplicate elements can be collected and rewritten as HadamardPower

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Hadamard_product_(matrices)
    """
    # Associativity
    rule = condition(
            lambda x: isinstance(x, HadamardProduct),
            flatten
        )
    fun = exhaust(rule)
    x = fun(x)

    # Identity
    fun = condition(
            lambda x: isinstance(x, HadamardProduct),
            rm_id(lambda x: isinstance(x, OneMatrix))
        )
    x = fun(x)

    # Absorbing by Zero Matrix
    def absorb(x):
        if any(isinstance(c, ZeroMatrix) for c in x.args):
            return ZeroMatrix(*x.shape)
        else:
            return x
    fun = condition(
            lambda x: isinstance(x, HadamardProduct),
            absorb
        )
    x = fun(x)

    # Rewriting with HadamardPower
    if isinstance(x, HadamardProduct):
        tally = Counter(x.args)

        new_arg = []
        for base, exp in tally.items():
            if exp == 1:
                new_arg.append(base)
            else:
                new_arg.append(HadamardPower(base, exp))

        x = HadamardProduct(*new_arg)

    # Commutativity
    fun = condition(
            lambda x: isinstance(x, HadamardProduct),
            sort(default_sort_key)
        )
    x = fun(x)

    # Unpacking
    x = unpack(x)
    return x


def hadamard_power(base, exp):
    base = sympify(base)
    exp = sympify(exp)
    if exp == 1:
        return base
    if not base.is_Matrix:
        return base**exp
    if exp.is_Matrix:
        raise ValueError("cannot raise expression to a matrix")
    return HadamardPower(base, exp)


class HadamardPower(MatrixExpr):
    r"""
    Elementwise power of matrix expressions

    Parameters
    ==========

    base : scalar or matrix

    exp : scalar or matrix

    Notes
    =====

    There are four definitions for the hadamard power which can be used.
    Let's consider `A, B` as `(m, n)` matrices, and `a, b` as scalars.

    Matrix raised to a scalar exponent:

    .. math::
        A^{\circ b} = \begin{bmatrix}
        A_{0, 0}^b   & A_{0, 1}^b   & \cdots & A_{0, n-1}^b   \\
        A_{1, 0}^b   & A_{1, 1}^b   & \cdots & A_{1, n-1}^b   \\
        \vdots       & \vdots       & \ddots & \vdots         \\
        A_{m-1, 0}^b & A_{m-1, 1}^b & \cdots & A_{m-1, n-1}^b
        \end{bmatrix}

    Scalar raised to a matrix exponent:

    .. math::
        a^{\circ B} = \begin{bmatrix}
        a^{B_{0, 0}}   & a^{B_{0, 1}}   & \cdots & a^{B_{0, n-1}}   \\
        a^{B_{1, 0}}   & a^{B_{1, 1}}   & \cdots & a^{B_{1, n-1}}   \\
        \vdots         & \vdots         & \ddots & \vdots           \\
        a^{B_{m-1, 0}} & a^{B_{m-1, 1}} & \cdots & a^{B_{m-1, n-1}}
        \end{bmatrix}

    Matrix raised to a matrix exponent:

    .. math::
        A^{\circ B} = \begin{bmatrix}
        A_{0, 0}^{B_{0, 0}}     & A_{0, 1}^{B_{0, 1}}     &
        \cdots & A_{0, n-1}^{B_{0, n-1}}     \\
        A_{1, 0}^{B_{1, 0}}     & A_{1, 1}^{B_{1, 1}}     &
        \cdots & A_{1, n-1}^{B_{1, n-1}}     \\
        \vdots                  & \vdots                  &
        \ddots & \vdots                      \\
        A_{m-1, 0}^{B_{m-1, 0}} & A_{m-1, 1}^{B_{m-1, 1}} &
        \cdots & A_{m-1, n-1}^{B_{m-1, n-1}}
        \end{bmatrix}

    Scalar raised to a scalar exponent:

    .. math::
        a^{\circ b} = a^b
    """

    def __new__(cls, base, exp):
        base = sympify(base)
        exp = sympify(exp)

        if base.is_scalar and exp.is_scalar:
            return base ** exp

        if isinstance(base, MatrixExpr) and isinstance(exp, MatrixExpr):
            validate(base, exp)

        obj = super().__new__(cls, base, exp)
        return obj

    @property
    def base(self):
        return self._args[0]

    @property
    def exp(self):
        return self._args[1]

    @property
    def shape(self):
        if self.base.is_Matrix:
            return self.base.shape
        return self.exp.shape

    def _entry(self, i, j, **kwargs):
        base = self.base
        exp = self.exp

        if base.is_Matrix:
            a = base._entry(i, j, **kwargs)
        elif base.is_scalar:
            a = base
        else:
            raise ValueError(
                'The base {} must be a scalar or a matrix.'.format(base))

        if exp.is_Matrix:
            b = exp._entry(i, j, **kwargs)
        elif exp.is_scalar:
            b = exp
        else:
            raise ValueError(
                'The exponent {} must be a scalar or a matrix.'.format(exp))

        return a ** b

    def _eval_transpose(self):
        from sympy.matrices.expressions.transpose import transpose
        return HadamardPower(transpose(self.base), self.exp)

    def _eval_derivative(self, x):
        dexp = self.exp.diff(x)
        logbase = self.base.applyfunc(log)
        dlbase = logbase.diff(x)
        return hadamard_product(
            dexp*logbase + self.exp*dlbase,
            self
        )

    def _eval_derivative_matrix_lines(self, x):
        from sympy.tensor.array.expressions.array_expressions import ArrayTensorProduct
        from sympy.tensor.array.expressions.array_expressions import ArrayDiagonal
        from sympy.matrices.expressions.matexpr import _make_matrix

        lr = self.base._eval_derivative_matrix_lines(x)
        for i in lr:
            diagonal = [(1, 2), (3, 4)]
            diagonal = [e for j, e in enumerate(diagonal) if self.base.shape[j] != 1]
            l1 = i._lines[i._first_line_index]
            l2 = i._lines[i._second_line_index]
            subexpr = ExprBuilder(
                ArrayDiagonal,
                [
                    ExprBuilder(
                        ArrayTensorProduct,
                        [
                            ExprBuilder(_make_matrix, [l1]),
                            self.exp*hadamard_power(self.base, self.exp-1),
                            ExprBuilder(_make_matrix, [l2]),
                        ]
                    ),
                *diagonal],
                validator=ArrayDiagonal._validate
            )
            i._first_pointer_parent = subexpr.args[0].args[0].args
            i._first_pointer_index = 0
            i._first_line_index = 0
            i._second_pointer_parent = subexpr.args[0].args[2].args
            i._second_pointer_index = 0
            i._second_line_index = 0
            i._lines = [subexpr]
        return lr

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\plotting\pygletplot\plot.py ===
from threading import RLock

# it is sufficient to import "pyglet" here once
try:
    import pyglet.gl as pgl
except ImportError:
    raise ImportError("pyglet is required for plotting.\n "
                      "visit https://pyglet.org/")

from sympy.core.numbers import Integer
from sympy.external.gmpy import SYMPY_INTS
from sympy.geometry.entity import GeometryEntity
from sympy.plotting.pygletplot.plot_axes import PlotAxes
from sympy.plotting.pygletplot.plot_mode import PlotMode
from sympy.plotting.pygletplot.plot_object import PlotObject
from sympy.plotting.pygletplot.plot_window import PlotWindow
from sympy.plotting.pygletplot.util import parse_option_string
from sympy.utilities.decorator import doctest_depends_on
from sympy.utilities.iterables import is_sequence

from time import sleep
from os import getcwd, listdir

import ctypes

@doctest_depends_on(modules=('pyglet',))
class PygletPlot:
    """
    Plot Examples
    =============

    See examples/advanced/pyglet_plotting.py for many more examples.

    >>> from sympy.plotting.pygletplot import PygletPlot as Plot
    >>> from sympy.abc import x, y, z

    >>> Plot(x*y**3-y*x**3)
    [0]: -x**3*y + x*y**3, 'mode=cartesian'

    >>> p = Plot()
    >>> p[1] = x*y
    >>> p[1].color = z, (0.4,0.4,0.9), (0.9,0.4,0.4)

    >>> p = Plot()
    >>> p[1] =  x**2+y**2
    >>> p[2] = -x**2-y**2


    Variable Intervals
    ==================

    The basic format is [var, min, max, steps], but the
    syntax is flexible and arguments left out are taken
    from the defaults for the current coordinate mode:

    >>> Plot(x**2) # implies [x,-5,5,100]
    [0]: x**2, 'mode=cartesian'
    >>> Plot(x**2, [], []) # [x,-1,1,40], [y,-1,1,40]
    [0]: x**2, 'mode=cartesian'
    >>> Plot(x**2-y**2, [100], [100]) # [x,-1,1,100], [y,-1,1,100]
    [0]: x**2 - y**2, 'mode=cartesian'
    >>> Plot(x**2, [x,-13,13,100])
    [0]: x**2, 'mode=cartesian'
    >>> Plot(x**2, [-13,13]) # [x,-13,13,100]
    [0]: x**2, 'mode=cartesian'
    >>> Plot(x**2, [x,-13,13]) # [x,-13,13,10]
    [0]: x**2, 'mode=cartesian'
    >>> Plot(1*x, [], [x], mode='cylindrical')
    ... # [unbound_theta,0,2*Pi,40], [x,-1,1,20]
    [0]: x, 'mode=cartesian'


    Coordinate Modes
    ================

    Plot supports several curvilinear coordinate modes, and
    they independent for each plotted function. You can specify
    a coordinate mode explicitly with the 'mode' named argument,
    but it can be automatically determined for Cartesian or
    parametric plots, and therefore must only be specified for
    polar, cylindrical, and spherical modes.

    Specifically, Plot(function arguments) and Plot[n] =
    (function arguments) will interpret your arguments as a
    Cartesian plot if you provide one function and a parametric
    plot if you provide two or three functions. Similarly, the
    arguments will be interpreted as a curve if one variable is
    used, and a surface if two are used.

    Supported mode names by number of variables:

    1: parametric, cartesian, polar
    2: parametric, cartesian, cylindrical = polar, spherical

    >>> Plot(1, mode='spherical')


    Calculator-like Interface
    =========================

    >>> p = Plot(visible=False)
    >>> f = x**2
    >>> p[1] = f
    >>> p[2] = f.diff(x)
    >>> p[3] = f.diff(x).diff(x)
    >>> p
    [1]: x**2, 'mode=cartesian'
    [2]: 2*x, 'mode=cartesian'
    [3]: 2, 'mode=cartesian'
    >>> p.show()
    >>> p.clear()
    >>> p
    <blank plot>
    >>> p[1] =  x**2+y**2
    >>> p[1].style = 'solid'
    >>> p[2] = -x**2-y**2
    >>> p[2].style = 'wireframe'
    >>> p[1].color = z, (0.4,0.4,0.9), (0.9,0.4,0.4)
    >>> p[1].style = 'both'
    >>> p[2].style = 'both'
    >>> p.close()


    Plot Window Keyboard Controls
    =============================

    Screen Rotation:
        X,Y axis      Arrow Keys, A,S,D,W, Numpad 4,6,8,2
        Z axis        Q,E, Numpad 7,9

    Model Rotation:
        Z axis        Z,C, Numpad 1,3

    Zoom:             R,F, PgUp,PgDn, Numpad +,-

    Reset Camera:     X, Numpad 5

    Camera Presets:
        XY            F1
        XZ            F2
        YZ            F3
        Perspective   F4

    Sensitivity Modifier: SHIFT

    Axes Toggle:
        Visible       F5
        Colors        F6

    Close Window:     ESCAPE

    =============================

    """

    @doctest_depends_on(modules=('pyglet',))
    def __init__(self, *fargs, **win_args):
        """
        Positional Arguments
        ====================

        Any given positional arguments are used to
        initialize a plot function at index 1. In
        other words...

        >>> from sympy.plotting.pygletplot import PygletPlot as Plot
        >>> from sympy.abc import x
        >>> p = Plot(x**2, visible=False)

        ...is equivalent to...

        >>> p = Plot(visible=False)
        >>> p[1] = x**2

        Note that in earlier versions of the plotting
        module, you were able to specify multiple
        functions in the initializer. This functionality
        has been dropped in favor of better automatic
        plot plot_mode detection.


        Named Arguments
        ===============

        axes
            An option string of the form
            "key1=value1; key2 = value2" which
            can use the following options:

            style = ordinate
                none OR frame OR box OR ordinate

            stride = 0.25
                val OR (val_x, val_y, val_z)

            overlay = True (draw on top of plot)
                True OR False

            colored = False (False uses Black,
                             True uses colors
                             R,G,B = X,Y,Z)
                True OR False

            label_axes = False (display axis names
                                at endpoints)
                True OR False

        visible = True (show immediately
            True OR False


        The following named arguments are passed as
        arguments to window initialization:

        antialiasing = True
            True OR False

        ortho = False
            True OR False

        invert_mouse_zoom = False
            True OR False

        """
        # Register the plot modes
        from . import plot_modes # noqa

        self._win_args = win_args
        self._window = None

        self._render_lock = RLock()

        self._functions = {}
        self._pobjects = []
        self._screenshot = ScreenShot(self)

        axe_options = parse_option_string(win_args.pop('axes', ''))
        self.axes = PlotAxes(**axe_options)
        self._pobjects.append(self.axes)

        self[0] = fargs
        if win_args.get('visible', True):
            self.show()

    ## Window Interfaces

    def show(self):
        """
        Creates and displays a plot window, or activates it
        (gives it focus) if it has already been created.
        """
        if self._window and not self._window.has_exit:
            self._window.activate()
        else:
            self._win_args['visible'] = True
            self.axes.reset_resources()

            #if hasattr(self, '_doctest_depends_on'):
            #    self._win_args['runfromdoctester'] = True

            self._window = PlotWindow(self, **self._win_args)

    def close(self):
        """
        Closes the plot window.
        """
        if self._window:
            self._window.close()

    def saveimage(self, outfile=None, format='', size=(600, 500)):
        """
        Saves a screen capture of the plot window to an
        image file.

        If outfile is given, it can either be a path
        or a file object. Otherwise a png image will
        be saved to the current working directory.
        If the format is omitted, it is determined from
        the filename extension.
        """
        self._screenshot.save(outfile, format, size)

    ## Function List Interfaces

    def clear(self):
        """
        Clears the function list of this plot.
        """
        self._render_lock.acquire()
        self._functions = {}
        self.adjust_all_bounds()
        self._render_lock.release()

    def __getitem__(self, i):
        """
        Returns the function at position i in the
        function list.
        """
        return self._functions[i]

    def __setitem__(self, i, args):
        """
        Parses and adds a PlotMode to the function
        list.
        """
        if not (isinstance(i, (SYMPY_INTS, Integer)) and i >= 0):
            raise ValueError("Function index must "
                             "be an integer >= 0.")

        if isinstance(args, PlotObject):
            f = args
        else:
            if (not is_sequence(args)) or isinstance(args, GeometryEntity):
                args = [args]
            if len(args) == 0:
                return  # no arguments given
            kwargs = {"bounds_callback": self.adjust_all_bounds}
            f = PlotMode(*args, **kwargs)

        if f:
            self._render_lock.acquire()
            self._functions[i] = f
            self._render_lock.release()
        else:
            raise ValueError("Failed to parse '%s'."
                    % ', '.join(str(a) for a in args))

    def __delitem__(self, i):
        """
        Removes the function in the function list at
        position i.
        """
        self._render_lock.acquire()
        del self._functions[i]
        self.adjust_all_bounds()
        self._render_lock.release()

    def firstavailableindex(self):
        """
        Returns the first unused index in the function list.
        """
        i = 0
        self._render_lock.acquire()
        while i in self._functions:
            i += 1
        self._render_lock.release()
        return i

    def append(self, *args):
        """
        Parses and adds a PlotMode to the function
        list at the first available index.
        """
        self.__setitem__(self.firstavailableindex(), args)

    def __len__(self):
        """
        Returns the number of functions in the function list.
        """
        return len(self._functions)

    def __iter__(self):
        """
        Allows iteration of the function list.
        """
        return self._functions.itervalues()

    def __repr__(self):
        return str(self)

    def __str__(self):
        """
        Returns a string containing a new-line separated
        list of the functions in the function list.
        """
        s = ""
        if len(self._functions) == 0:
            s += "<blank plot>"
        else:
            self._render_lock.acquire()
            s += "\n".join(["%s[%i]: %s" % ("", i, str(self._functions[i]))
                            for i in self._functions])
            self._render_lock.release()
        return s

    def adjust_all_bounds(self):
        self._render_lock.acquire()
        self.axes.reset_bounding_box()
        for f in self._functions:
            self.axes.adjust_bounds(self._functions[f].bounds)
        self._render_lock.release()

    def wait_for_calculations(self):
        sleep(0)
        self._render_lock.acquire()
        for f in self._functions:
            a = self._functions[f]._get_calculating_verts
            b = self._functions[f]._get_calculating_cverts
            while a() or b():
                sleep(0)
        self._render_lock.release()

class ScreenShot:
    def __init__(self, plot):
        self._plot = plot
        self.screenshot_requested = False
        self.outfile = None
        self.format = ''
        self.invisibleMode = False
        self.flag = 0

    def __bool__(self):
        return self.screenshot_requested

    def _execute_saving(self):
        if self.flag < 3:
            self.flag += 1
            return

        size_x, size_y = self._plot._window.get_size()
        size = size_x*size_y*4*ctypes.sizeof(ctypes.c_ubyte)
        image = ctypes.create_string_buffer(size)
        pgl.glReadPixels(0, 0, size_x, size_y, pgl.GL_RGBA, pgl.GL_UNSIGNED_BYTE, image)
        from PIL import Image
        im = Image.frombuffer('RGBA', (size_x, size_y),
                              image.raw, 'raw', 'RGBA', 0, 1)
        im.transpose(Image.FLIP_TOP_BOTTOM).save(self.outfile, self.format)

        self.flag = 0
        self.screenshot_requested = False
        if self.invisibleMode:
            self._plot._window.close()

    def save(self, outfile=None, format='', size=(600, 500)):
        self.outfile = outfile
        self.format = format
        self.size = size
        self.screenshot_requested = True

        if not self._plot._window or self._plot._window.has_exit:
            self._plot._win_args['visible'] = False

            self._plot._win_args['width'] = size[0]
            self._plot._win_args['height'] = size[1]

            self._plot.axes.reset_resources()
            self._plot._window = PlotWindow(self._plot, **self._plot._win_args)
            self.invisibleMode = True

        if self.outfile is None:
            self.outfile = self._create_unique_path()
            print(self.outfile)

    def _create_unique_path(self):
        cwd = getcwd()
        l = listdir(cwd)
        path = ''
        i = 0
        while True:
            if not 'plot_%s.png' % i in l:
                path = cwd + '/plot_%s.png' % i
                break
            i += 1
        return path

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\interactive\session.py ===
"""Tools for setting up interactive sessions. """

from sympy.external.gmpy import GROUND_TYPES
from sympy.external.importtools import version_tuple

from sympy.interactive.printing import init_printing

from sympy.utilities.misc import ARCH

preexec_source = """\
from sympy import *
x, y, z, t = symbols('x y z t')
k, m, n = symbols('k m n', integer=True)
f, g, h = symbols('f g h', cls=Function)
init_printing()
"""

verbose_message = """\
These commands were executed:
%(source)s
Documentation can be found at https://docs.sympy.org/%(version)s
"""

no_ipython = """\
Could not locate IPython. Having IPython installed is greatly recommended.
See http://ipython.scipy.org for more details. If you use Debian/Ubuntu,
just install the 'ipython' package and start isympy again.
"""


def _make_message(ipython=True, quiet=False, source=None):
    """Create a banner for an interactive session. """
    from sympy import __version__ as sympy_version
    from sympy import SYMPY_DEBUG

    import sys
    import os

    if quiet:
        return ""

    python_version = "%d.%d.%d" % sys.version_info[:3]

    if ipython:
        shell_name = "IPython"
    else:
        shell_name = "Python"

    info = ['ground types: %s' % GROUND_TYPES]

    cache = os.getenv('SYMPY_USE_CACHE')

    if cache is not None and cache.lower() == 'no':
        info.append('cache: off')

    if SYMPY_DEBUG:
        info.append('debugging: on')

    args = shell_name, sympy_version, python_version, ARCH, ', '.join(info)
    message = "%s console for SymPy %s (Python %s-%s) (%s)\n" % args

    if source is None:
        source = preexec_source

    _source = ""

    for line in source.split('\n')[:-1]:
        if not line:
            _source += '\n'
        else:
            _source += '>>> ' + line + '\n'

    doc_version = sympy_version
    if 'dev' in doc_version:
        doc_version = "dev"
    else:
        doc_version = "%s/" % doc_version

    message += '\n' + verbose_message % {'source': _source,
                                         'version': doc_version}

    return message


def int_to_Integer(s):
    """
    Wrap integer literals with Integer.

    This is based on the decistmt example from
    https://docs.python.org/3/library/tokenize.html.

    Only integer literals are converted.  Float literals are left alone.

    Examples
    ========

    >>> from sympy import Integer # noqa: F401
    >>> from sympy.interactive.session import int_to_Integer
    >>> s = '1.2 + 1/2 - 0x12 + a1'
    >>> int_to_Integer(s)
    '1.2 +Integer (1 )/Integer (2 )-Integer (0x12 )+a1 '
    >>> s = 'print (1/2)'
    >>> int_to_Integer(s)
    'print (Integer (1 )/Integer (2 ))'
    >>> exec(s)
    0.5
    >>> exec(int_to_Integer(s))
    1/2
    """
    from tokenize import generate_tokens, untokenize, NUMBER, NAME, OP
    from io import StringIO

    def _is_int(num):
        """
        Returns true if string value num (with token NUMBER) represents an integer.
        """
        # XXX: Is there something in the standard library that will do this?
        if '.' in num or 'j' in num.lower() or 'e' in num.lower():
            return False
        return True

    result = []
    g = generate_tokens(StringIO(s).readline)  # tokenize the string
    for toknum, tokval, _, _, _ in g:
        if toknum == NUMBER and _is_int(tokval):  # replace NUMBER tokens
            result.extend([
                (NAME, 'Integer'),
                (OP, '('),
                (NUMBER, tokval),
                (OP, ')')
            ])
        else:
            result.append((toknum, tokval))
    return untokenize(result)


def enable_automatic_int_sympification(shell):
    """
    Allow IPython to automatically convert integer literals to Integer.
    """
    import ast
    old_run_cell = shell.run_cell

    def my_run_cell(cell, *args, **kwargs):
        try:
            # Check the cell for syntax errors.  This way, the syntax error
            # will show the original input, not the transformed input.  The
            # downside here is that IPython magic like %timeit will not work
            # with transformed input (but on the other hand, IPython magic
            # that doesn't expect transformed input will continue to work).
            ast.parse(cell)
        except SyntaxError:
            pass
        else:
            cell = int_to_Integer(cell)
        return old_run_cell(cell, *args, **kwargs)

    shell.run_cell = my_run_cell


def enable_automatic_symbols(shell):
    """Allow IPython to automatically create symbols (``isympy -a``). """
    # XXX: This should perhaps use tokenize, like int_to_Integer() above.
    # This would avoid re-executing the code, which can lead to subtle
    # issues.  For example:
    #
    # In [1]: a = 1
    #
    # In [2]: for i in range(10):
    #    ...:     a += 1
    #    ...:
    #
    # In [3]: a
    # Out[3]: 11
    #
    # In [4]: a = 1
    #
    # In [5]: for i in range(10):
    #    ...:     a += 1
    #    ...:     print b
    #    ...:
    # b
    # b
    # b
    # b
    # b
    # b
    # b
    # b
    # b
    # b
    #
    # In [6]: a
    # Out[6]: 12
    #
    # Note how the for loop is executed again because `b` was not defined, but `a`
    # was already incremented once, so the result is that it is incremented
    # multiple times.

    import re
    re_nameerror = re.compile(
        "name '(?P<symbol>[A-Za-z_][A-Za-z0-9_]*)' is not defined")

    def _handler(self, etype, value, tb, tb_offset=None):
        """Handle :exc:`NameError` exception and allow injection of missing symbols. """
        if etype is NameError and tb.tb_next and not tb.tb_next.tb_next:
            match = re_nameerror.match(str(value))

            if match is not None:
                # XXX: Make sure Symbol is in scope. Otherwise you'll get infinite recursion.
                self.run_cell("%(symbol)s = Symbol('%(symbol)s')" %
                              {'symbol': match.group("symbol")}, store_history=False)

                try:
                    code = self.user_ns['In'][-1]
                except (KeyError, IndexError):
                    pass
                else:
                    self.run_cell(code, store_history=False)
                    return None
                finally:
                    self.run_cell("del %s" % match.group("symbol"),
                                  store_history=False)

        stb = self.InteractiveTB.structured_traceback(
            etype, value, tb, tb_offset=tb_offset)
        self._showtraceback(etype, value, stb)

    shell.set_custom_exc((NameError,), _handler)


def init_ipython_session(shell=None, argv=[], auto_symbols=False, auto_int_to_Integer=False):
    """Construct new IPython session. """
    import IPython

    if version_tuple(IPython.__version__) >= version_tuple('0.11'):
        if not shell:
            # use an app to parse the command line, and init config
            # IPython 1.0 deprecates the frontend module, so we import directly
            # from the terminal module to prevent a deprecation message from being
            # shown.
            if version_tuple(IPython.__version__) >= version_tuple('1.0'):
                from IPython.terminal import ipapp
            else:
                from IPython.frontend.terminal import ipapp
            app = ipapp.TerminalIPythonApp()

            # don't draw IPython banner during initialization:
            app.display_banner = False
            app.initialize(argv)

            shell = app.shell

        if auto_symbols:
            enable_automatic_symbols(shell)
        if auto_int_to_Integer:
            enable_automatic_int_sympification(shell)

        return shell
    else:
        from IPython.Shell import make_IPython
        return make_IPython(argv)


def init_python_session():
    """Construct new Python session. """
    from code import InteractiveConsole

    class SymPyConsole(InteractiveConsole):
        """An interactive console with readline support. """

        def __init__(self):
            ns_locals = {}
            InteractiveConsole.__init__(self, locals=ns_locals)
            try:
                import rlcompleter
                import readline
            except ImportError:
                pass
            else:
                import os
                import atexit

                readline.set_completer(rlcompleter.Completer(ns_locals).complete)
                readline.parse_and_bind('tab: complete')

                if hasattr(readline, 'read_history_file'):
                    history = os.path.expanduser('~/.sympy-history')

                    try:
                        readline.read_history_file(history)
                    except OSError:
                        pass

                    atexit.register(readline.write_history_file, history)

    return SymPyConsole()


def init_session(ipython=None, pretty_print=True, order=None,
                 use_unicode=None, use_latex=None, quiet=False, auto_symbols=False,
                 auto_int_to_Integer=False, str_printer=None, pretty_printer=None,
                 latex_printer=None, argv=[]):
    """
    Initialize an embedded IPython or Python session. The IPython session is
    initiated with the --pylab option, without the numpy imports, so that
    matplotlib plotting can be interactive.

    Parameters
    ==========

    pretty_print: boolean
        If True, use pretty_print to stringify;
        if False, use sstrrepr to stringify.
    order: string or None
        There are a few different settings for this parameter:
        lex (default), which is lexographic order;
        grlex, which is graded lexographic order;
        grevlex, which is reversed graded lexographic order;
        old, which is used for compatibility reasons and for long expressions;
        None, which sets it to lex.
    use_unicode: boolean or None
        If True, use unicode characters;
        if False, do not use unicode characters.
    use_latex: boolean or None
        If True, use latex rendering if IPython GUI's;
        if False, do not use latex rendering.
    quiet: boolean
        If True, init_session will not print messages regarding its status;
        if False, init_session will print messages regarding its status.
    auto_symbols: boolean
        If True, IPython will automatically create symbols for you.
        If False, it will not.
        The default is False.
    auto_int_to_Integer: boolean
        If True, IPython will automatically wrap int literals with Integer, so
        that things like 1/2 give Rational(1, 2).
        If False, it will not.
        The default is False.
    ipython: boolean or None
        If True, printing will initialize for an IPython console;
        if False, printing will initialize for a normal console;
        The default is None, which automatically determines whether we are in
        an ipython instance or not.
    str_printer: function, optional, default=None
        A custom string printer function. This should mimic
        sympy.printing.sstrrepr().
    pretty_printer: function, optional, default=None
        A custom pretty printer. This should mimic sympy.printing.pretty().
    latex_printer: function, optional, default=None
        A custom LaTeX printer. This should mimic sympy.printing.latex()
        This should mimic sympy.printing.latex().
    argv: list of arguments for IPython
        See sympy.bin.isympy for options that can be used to initialize IPython.

    See Also
    ========

    sympy.interactive.printing.init_printing: for examples and the rest of the parameters.


    Examples
    ========

    >>> from sympy import init_session, Symbol, sin, sqrt
    >>> sin(x) #doctest: +SKIP
    NameError: name 'x' is not defined
    >>> init_session() #doctest: +SKIP
    >>> sin(x) #doctest: +SKIP
    sin(x)
    >>> sqrt(5) #doctest: +SKIP
      ___
    \\/ 5
    >>> init_session(pretty_print=False) #doctest: +SKIP
    >>> sqrt(5) #doctest: +SKIP
    sqrt(5)
    >>> y + x + y**2 + x**2 #doctest: +SKIP
    x**2 + x + y**2 + y
    >>> init_session(order='grlex') #doctest: +SKIP
    >>> y + x + y**2 + x**2 #doctest: +SKIP
    x**2 + y**2 + x + y
    >>> init_session(order='grevlex') #doctest: +SKIP
    >>> y * x**2 + x * y**2 #doctest: +SKIP
    x**2*y + x*y**2
    >>> init_session(order='old') #doctest: +SKIP
    >>> x**2 + y**2 + x + y #doctest: +SKIP
    x + y + x**2 + y**2
    >>> theta = Symbol('theta') #doctest: +SKIP
    >>> theta #doctest: +SKIP
    theta
    >>> init_session(use_unicode=True) #doctest: +SKIP
    >>> theta # doctest: +SKIP
    \u03b8
    """
    import sys

    in_ipython = False

    if ipython is not False:
        try:
            import IPython
        except ImportError:
            if ipython is True:
                raise RuntimeError("IPython is not available on this system")
            ip = None
        else:
            try:
                from IPython import get_ipython
                ip = get_ipython()
            except ImportError:
                ip = None
        in_ipython = bool(ip)
        if ipython is None:
            ipython = in_ipython

    if ipython is False:
        ip = init_python_session()
        mainloop = ip.interact
    else:
        ip = init_ipython_session(ip, argv=argv, auto_symbols=auto_symbols,
                                  auto_int_to_Integer=auto_int_to_Integer)

        if version_tuple(IPython.__version__) >= version_tuple('0.11'):
            # runsource is gone, use run_cell instead, which doesn't
            # take a symbol arg.  The second arg is `store_history`,
            # and False means don't add the line to IPython's history.
            ip.runsource = lambda src, symbol='exec': ip.run_cell(src, False)

            # Enable interactive plotting using pylab.
            try:
                ip.enable_pylab(import_all=False)
            except Exception:
                # Causes an import error if matplotlib is not installed.
                # Causes other errors (depending on the backend) if there
                # is no display, or if there is some problem in the
                # backend, so we have a bare "except Exception" here
                pass
        if not in_ipython:
            mainloop = ip.mainloop

    if auto_symbols and (not ipython or version_tuple(IPython.__version__) < version_tuple('0.11')):
        raise RuntimeError("automatic construction of symbols is possible only in IPython 0.11 or above")
    if auto_int_to_Integer and (not ipython or version_tuple(IPython.__version__) < version_tuple('0.11')):
        raise RuntimeError("automatic int to Integer transformation is possible only in IPython 0.11 or above")

    _preexec_source = preexec_source

    ip.runsource(_preexec_source, symbol='exec')
    init_printing(pretty_print=pretty_print, order=order,
                  use_unicode=use_unicode, use_latex=use_latex, ip=ip,
                  str_printer=str_printer, pretty_printer=pretty_printer,
                  latex_printer=latex_printer)

    message = _make_message(ipython, quiet, _preexec_source)

    if not in_ipython:
        print(message)
        mainloop()
        sys.exit('Exiting ...')
    else:
        print(message)
        import atexit
        atexit.register(lambda: print("Exiting ...\n"))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v135\layer_tree.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: LayerTree (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing
from . import dom


class LayerId(str):
    '''
    Unique Layer identifier.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> LayerId:
        return cls(json)

    def __repr__(self):
        return 'LayerId({})'.format(super().__repr__())


class SnapshotId(str):
    '''
    Unique snapshot identifier.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> SnapshotId:
        return cls(json)

    def __repr__(self):
        return 'SnapshotId({})'.format(super().__repr__())


@dataclass
class ScrollRect:
    '''
    Rectangle where scrolling happens on the main thread.
    '''
    #: Rectangle itself.
    rect: dom.Rect

    #: Reason for rectangle to force scrolling on the main thread
    type_: str

    def to_json(self):
        json = dict()
        json['rect'] = self.rect.to_json()
        json['type'] = self.type_
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            rect=dom.Rect.from_json(json['rect']),
            type_=str(json['type']),
        )


@dataclass
class StickyPositionConstraint:
    '''
    Sticky position constraints.
    '''
    #: Layout rectangle of the sticky element before being shifted
    sticky_box_rect: dom.Rect

    #: Layout rectangle of the containing block of the sticky element
    containing_block_rect: dom.Rect

    #: The nearest sticky layer that shifts the sticky box
    nearest_layer_shifting_sticky_box: typing.Optional[LayerId] = None

    #: The nearest sticky layer that shifts the containing block
    nearest_layer_shifting_containing_block: typing.Optional[LayerId] = None

    def to_json(self):
        json = dict()
        json['stickyBoxRect'] = self.sticky_box_rect.to_json()
        json['containingBlockRect'] = self.containing_block_rect.to_json()
        if self.nearest_layer_shifting_sticky_box is not None:
            json['nearestLayerShiftingStickyBox'] = self.nearest_layer_shifting_sticky_box.to_json()
        if self.nearest_layer_shifting_containing_block is not None:
            json['nearestLayerShiftingContainingBlock'] = self.nearest_layer_shifting_containing_block.to_json()
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            sticky_box_rect=dom.Rect.from_json(json['stickyBoxRect']),
            containing_block_rect=dom.Rect.from_json(json['containingBlockRect']),
            nearest_layer_shifting_sticky_box=LayerId.from_json(json['nearestLayerShiftingStickyBox']) if 'nearestLayerShiftingStickyBox' in json else None,
            nearest_layer_shifting_containing_block=LayerId.from_json(json['nearestLayerShiftingContainingBlock']) if 'nearestLayerShiftingContainingBlock' in json else None,
        )


@dataclass
class PictureTile:
    '''
    Serialized fragment of layer picture along with its offset within the layer.
    '''
    #: Offset from owning layer left boundary
    x: float

    #: Offset from owning layer top boundary
    y: float

    #: Base64-encoded snapshot data.
    picture: str

    def to_json(self):
        json = dict()
        json['x'] = self.x
        json['y'] = self.y
        json['picture'] = self.picture
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            x=float(json['x']),
            y=float(json['y']),
            picture=str(json['picture']),
        )


@dataclass
class Layer:
    '''
    Information about a compositing layer.
    '''
    #: The unique id for this layer.
    layer_id: LayerId

    #: Offset from parent layer, X coordinate.
    offset_x: float

    #: Offset from parent layer, Y coordinate.
    offset_y: float

    #: Layer width.
    width: float

    #: Layer height.
    height: float

    #: Indicates how many time this layer has painted.
    paint_count: int

    #: Indicates whether this layer hosts any content, rather than being used for
    #: transform/scrolling purposes only.
    draws_content: bool

    #: The id of parent (not present for root).
    parent_layer_id: typing.Optional[LayerId] = None

    #: The backend id for the node associated with this layer.
    backend_node_id: typing.Optional[dom.BackendNodeId] = None

    #: Transformation matrix for layer, default is identity matrix
    transform: typing.Optional[typing.List[float]] = None

    #: Transform anchor point X, absent if no transform specified
    anchor_x: typing.Optional[float] = None

    #: Transform anchor point Y, absent if no transform specified
    anchor_y: typing.Optional[float] = None

    #: Transform anchor point Z, absent if no transform specified
    anchor_z: typing.Optional[float] = None

    #: Set if layer is not visible.
    invisible: typing.Optional[bool] = None

    #: Rectangles scrolling on main thread only.
    scroll_rects: typing.Optional[typing.List[ScrollRect]] = None

    #: Sticky position constraint information
    sticky_position_constraint: typing.Optional[StickyPositionConstraint] = None

    def to_json(self):
        json = dict()
        json['layerId'] = self.layer_id.to_json()
        json['offsetX'] = self.offset_x
        json['offsetY'] = self.offset_y
        json['width'] = self.width
        json['height'] = self.height
        json['paintCount'] = self.paint_count
        json['drawsContent'] = self.draws_content
        if self.parent_layer_id is not None:
            json['parentLayerId'] = self.parent_layer_id.to_json()
        if self.backend_node_id is not None:
            json['backendNodeId'] = self.backend_node_id.to_json()
        if self.transform is not None:
            json['transform'] = [i for i in self.transform]
        if self.anchor_x is not None:
            json['anchorX'] = self.anchor_x
        if self.anchor_y is not None:
            json['anchorY'] = self.anchor_y
        if self.anchor_z is not None:
            json['anchorZ'] = self.anchor_z
        if self.invisible is not None:
            json['invisible'] = self.invisible
        if self.scroll_rects is not None:
            json['scrollRects'] = [i.to_json() for i in self.scroll_rects]
        if self.sticky_position_constraint is not None:
            json['stickyPositionConstraint'] = self.sticky_position_constraint.to_json()
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            layer_id=LayerId.from_json(json['layerId']),
            offset_x=float(json['offsetX']),
            offset_y=float(json['offsetY']),
            width=float(json['width']),
            height=float(json['height']),
            paint_count=int(json['paintCount']),
            draws_content=bool(json['drawsContent']),
            parent_layer_id=LayerId.from_json(json['parentLayerId']) if 'parentLayerId' in json else None,
            backend_node_id=dom.BackendNodeId.from_json(json['backendNodeId']) if 'backendNodeId' in json else None,
            transform=[float(i) for i in json['transform']] if 'transform' in json else None,
            anchor_x=float(json['anchorX']) if 'anchorX' in json else None,
            anchor_y=float(json['anchorY']) if 'anchorY' in json else None,
            anchor_z=float(json['anchorZ']) if 'anchorZ' in json else None,
            invisible=bool(json['invisible']) if 'invisible' in json else None,
            scroll_rects=[ScrollRect.from_json(i) for i in json['scrollRects']] if 'scrollRects' in json else None,
            sticky_position_constraint=StickyPositionConstraint.from_json(json['stickyPositionConstraint']) if 'stickyPositionConstraint' in json else None,
        )


class PaintProfile(list):
    '''
    Array of timings, one per paint step.
    '''
    def to_json(self) -> typing.List[float]:
        return self

    @classmethod
    def from_json(cls, json: typing.List[float]) -> PaintProfile:
        return cls(json)

    def __repr__(self):
        return 'PaintProfile({})'.format(super().__repr__())


def compositing_reasons(
        layer_id: LayerId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.Tuple[typing.List[str], typing.List[str]]]:
    '''
    Provides the reasons why the given layer was composited.

    :param layer_id: The id of the layer for which we want to get the reasons it was composited.
    :returns: A tuple with the following items:

        0. **compositingReasons** - A list of strings specifying reasons for the given layer to become composited.
        1. **compositingReasonIds** - A list of strings specifying reason IDs for the given layer to become composited.
    '''
    params: T_JSON_DICT = dict()
    params['layerId'] = layer_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.compositingReasons',
        'params': params,
    }
    json = yield cmd_dict
    return (
        [str(i) for i in json['compositingReasons']],
        [str(i) for i in json['compositingReasonIds']]
    )


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disables compositing tree inspection.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.disable',
    }
    json = yield cmd_dict


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enables compositing tree inspection.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.enable',
    }
    json = yield cmd_dict


def load_snapshot(
        tiles: typing.List[PictureTile]
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SnapshotId]:
    '''
    Returns the snapshot identifier.

    :param tiles: An array of tiles composing the snapshot.
    :returns: The id of the snapshot.
    '''
    params: T_JSON_DICT = dict()
    params['tiles'] = [i.to_json() for i in tiles]
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.loadSnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return SnapshotId.from_json(json['snapshotId'])


def make_snapshot(
        layer_id: LayerId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SnapshotId]:
    '''
    Returns the layer snapshot identifier.

    :param layer_id: The id of the layer.
    :returns: The id of the layer snapshot.
    '''
    params: T_JSON_DICT = dict()
    params['layerId'] = layer_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.makeSnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return SnapshotId.from_json(json['snapshotId'])


def profile_snapshot(
        snapshot_id: SnapshotId,
        min_repeat_count: typing.Optional[int] = None,
        min_duration: typing.Optional[float] = None,
        clip_rect: typing.Optional[dom.Rect] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[PaintProfile]]:
    '''
    :param snapshot_id: The id of the layer snapshot.
    :param min_repeat_count: *(Optional)* The maximum number of times to replay the snapshot (1, if not specified).
    :param min_duration: *(Optional)* The minimum duration (in seconds) to replay the snapshot.
    :param clip_rect: *(Optional)* The clip rectangle to apply when replaying the snapshot.
    :returns: The array of paint profiles, one per run.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    if min_repeat_count is not None:
        params['minRepeatCount'] = min_repeat_count
    if min_duration is not None:
        params['minDuration'] = min_duration
    if clip_rect is not None:
        params['clipRect'] = clip_rect.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.profileSnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return [PaintProfile.from_json(i) for i in json['timings']]


def release_snapshot(
        snapshot_id: SnapshotId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Releases layer snapshot captured by the back-end.

    :param snapshot_id: The id of the layer snapshot.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.releaseSnapshot',
        'params': params,
    }
    json = yield cmd_dict


def replay_snapshot(
        snapshot_id: SnapshotId,
        from_step: typing.Optional[int] = None,
        to_step: typing.Optional[int] = None,
        scale: typing.Optional[float] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,str]:
    '''
    Replays the layer snapshot and returns the resulting bitmap.

    :param snapshot_id: The id of the layer snapshot.
    :param from_step: *(Optional)* The first step to replay from (replay from the very start if not specified).
    :param to_step: *(Optional)* The last step to replay to (replay till the end if not specified).
    :param scale: *(Optional)* The scale to apply while replaying (defaults to 1).
    :returns: A data: URL for resulting image.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    if from_step is not None:
        params['fromStep'] = from_step
    if to_step is not None:
        params['toStep'] = to_step
    if scale is not None:
        params['scale'] = scale
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.replaySnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return str(json['dataURL'])


def snapshot_command_log(
        snapshot_id: SnapshotId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[dict]]:
    '''
    Replays the layer snapshot and returns canvas log.

    :param snapshot_id: The id of the layer snapshot.
    :returns: The array of canvas function calls.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.snapshotCommandLog',
        'params': params,
    }
    json = yield cmd_dict
    return [dict(i) for i in json['commandLog']]


@event_class('LayerTree.layerPainted')
@dataclass
class LayerPainted:
    #: The id of the painted layer.
    layer_id: LayerId
    #: Clip rectangle.
    clip: dom.Rect

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> LayerPainted:
        return cls(
            layer_id=LayerId.from_json(json['layerId']),
            clip=dom.Rect.from_json(json['clip'])
        )


@event_class('LayerTree.layerTreeDidChange')
@dataclass
class LayerTreeDidChange:
    #: Layer tree, absent if not in the compositing mode.
    layers: typing.Optional[typing.List[Layer]]

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> LayerTreeDidChange:
        return cls(
            layers=[Layer.from_json(i) for i in json['layers']] if 'layers' in json else None
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v136\layer_tree.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: LayerTree (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing
from . import dom


class LayerId(str):
    '''
    Unique Layer identifier.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> LayerId:
        return cls(json)

    def __repr__(self):
        return 'LayerId({})'.format(super().__repr__())


class SnapshotId(str):
    '''
    Unique snapshot identifier.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> SnapshotId:
        return cls(json)

    def __repr__(self):
        return 'SnapshotId({})'.format(super().__repr__())


@dataclass
class ScrollRect:
    '''
    Rectangle where scrolling happens on the main thread.
    '''
    #: Rectangle itself.
    rect: dom.Rect

    #: Reason for rectangle to force scrolling on the main thread
    type_: str

    def to_json(self):
        json = dict()
        json['rect'] = self.rect.to_json()
        json['type'] = self.type_
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            rect=dom.Rect.from_json(json['rect']),
            type_=str(json['type']),
        )


@dataclass
class StickyPositionConstraint:
    '''
    Sticky position constraints.
    '''
    #: Layout rectangle of the sticky element before being shifted
    sticky_box_rect: dom.Rect

    #: Layout rectangle of the containing block of the sticky element
    containing_block_rect: dom.Rect

    #: The nearest sticky layer that shifts the sticky box
    nearest_layer_shifting_sticky_box: typing.Optional[LayerId] = None

    #: The nearest sticky layer that shifts the containing block
    nearest_layer_shifting_containing_block: typing.Optional[LayerId] = None

    def to_json(self):
        json = dict()
        json['stickyBoxRect'] = self.sticky_box_rect.to_json()
        json['containingBlockRect'] = self.containing_block_rect.to_json()
        if self.nearest_layer_shifting_sticky_box is not None:
            json['nearestLayerShiftingStickyBox'] = self.nearest_layer_shifting_sticky_box.to_json()
        if self.nearest_layer_shifting_containing_block is not None:
            json['nearestLayerShiftingContainingBlock'] = self.nearest_layer_shifting_containing_block.to_json()
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            sticky_box_rect=dom.Rect.from_json(json['stickyBoxRect']),
            containing_block_rect=dom.Rect.from_json(json['containingBlockRect']),
            nearest_layer_shifting_sticky_box=LayerId.from_json(json['nearestLayerShiftingStickyBox']) if 'nearestLayerShiftingStickyBox' in json else None,
            nearest_layer_shifting_containing_block=LayerId.from_json(json['nearestLayerShiftingContainingBlock']) if 'nearestLayerShiftingContainingBlock' in json else None,
        )


@dataclass
class PictureTile:
    '''
    Serialized fragment of layer picture along with its offset within the layer.
    '''
    #: Offset from owning layer left boundary
    x: float

    #: Offset from owning layer top boundary
    y: float

    #: Base64-encoded snapshot data.
    picture: str

    def to_json(self):
        json = dict()
        json['x'] = self.x
        json['y'] = self.y
        json['picture'] = self.picture
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            x=float(json['x']),
            y=float(json['y']),
            picture=str(json['picture']),
        )


@dataclass
class Layer:
    '''
    Information about a compositing layer.
    '''
    #: The unique id for this layer.
    layer_id: LayerId

    #: Offset from parent layer, X coordinate.
    offset_x: float

    #: Offset from parent layer, Y coordinate.
    offset_y: float

    #: Layer width.
    width: float

    #: Layer height.
    height: float

    #: Indicates how many time this layer has painted.
    paint_count: int

    #: Indicates whether this layer hosts any content, rather than being used for
    #: transform/scrolling purposes only.
    draws_content: bool

    #: The id of parent (not present for root).
    parent_layer_id: typing.Optional[LayerId] = None

    #: The backend id for the node associated with this layer.
    backend_node_id: typing.Optional[dom.BackendNodeId] = None

    #: Transformation matrix for layer, default is identity matrix
    transform: typing.Optional[typing.List[float]] = None

    #: Transform anchor point X, absent if no transform specified
    anchor_x: typing.Optional[float] = None

    #: Transform anchor point Y, absent if no transform specified
    anchor_y: typing.Optional[float] = None

    #: Transform anchor point Z, absent if no transform specified
    anchor_z: typing.Optional[float] = None

    #: Set if layer is not visible.
    invisible: typing.Optional[bool] = None

    #: Rectangles scrolling on main thread only.
    scroll_rects: typing.Optional[typing.List[ScrollRect]] = None

    #: Sticky position constraint information
    sticky_position_constraint: typing.Optional[StickyPositionConstraint] = None

    def to_json(self):
        json = dict()
        json['layerId'] = self.layer_id.to_json()
        json['offsetX'] = self.offset_x
        json['offsetY'] = self.offset_y
        json['width'] = self.width
        json['height'] = self.height
        json['paintCount'] = self.paint_count
        json['drawsContent'] = self.draws_content
        if self.parent_layer_id is not None:
            json['parentLayerId'] = self.parent_layer_id.to_json()
        if self.backend_node_id is not None:
            json['backendNodeId'] = self.backend_node_id.to_json()
        if self.transform is not None:
            json['transform'] = [i for i in self.transform]
        if self.anchor_x is not None:
            json['anchorX'] = self.anchor_x
        if self.anchor_y is not None:
            json['anchorY'] = self.anchor_y
        if self.anchor_z is not None:
            json['anchorZ'] = self.anchor_z
        if self.invisible is not None:
            json['invisible'] = self.invisible
        if self.scroll_rects is not None:
            json['scrollRects'] = [i.to_json() for i in self.scroll_rects]
        if self.sticky_position_constraint is not None:
            json['stickyPositionConstraint'] = self.sticky_position_constraint.to_json()
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            layer_id=LayerId.from_json(json['layerId']),
            offset_x=float(json['offsetX']),
            offset_y=float(json['offsetY']),
            width=float(json['width']),
            height=float(json['height']),
            paint_count=int(json['paintCount']),
            draws_content=bool(json['drawsContent']),
            parent_layer_id=LayerId.from_json(json['parentLayerId']) if 'parentLayerId' in json else None,
            backend_node_id=dom.BackendNodeId.from_json(json['backendNodeId']) if 'backendNodeId' in json else None,
            transform=[float(i) for i in json['transform']] if 'transform' in json else None,
            anchor_x=float(json['anchorX']) if 'anchorX' in json else None,
            anchor_y=float(json['anchorY']) if 'anchorY' in json else None,
            anchor_z=float(json['anchorZ']) if 'anchorZ' in json else None,
            invisible=bool(json['invisible']) if 'invisible' in json else None,
            scroll_rects=[ScrollRect.from_json(i) for i in json['scrollRects']] if 'scrollRects' in json else None,
            sticky_position_constraint=StickyPositionConstraint.from_json(json['stickyPositionConstraint']) if 'stickyPositionConstraint' in json else None,
        )


class PaintProfile(list):
    '''
    Array of timings, one per paint step.
    '''
    def to_json(self) -> typing.List[float]:
        return self

    @classmethod
    def from_json(cls, json: typing.List[float]) -> PaintProfile:
        return cls(json)

    def __repr__(self):
        return 'PaintProfile({})'.format(super().__repr__())


def compositing_reasons(
        layer_id: LayerId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.Tuple[typing.List[str], typing.List[str]]]:
    '''
    Provides the reasons why the given layer was composited.

    :param layer_id: The id of the layer for which we want to get the reasons it was composited.
    :returns: A tuple with the following items:

        0. **compositingReasons** - A list of strings specifying reasons for the given layer to become composited.
        1. **compositingReasonIds** - A list of strings specifying reason IDs for the given layer to become composited.
    '''
    params: T_JSON_DICT = dict()
    params['layerId'] = layer_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.compositingReasons',
        'params': params,
    }
    json = yield cmd_dict
    return (
        [str(i) for i in json['compositingReasons']],
        [str(i) for i in json['compositingReasonIds']]
    )


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disables compositing tree inspection.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.disable',
    }
    json = yield cmd_dict


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enables compositing tree inspection.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.enable',
    }
    json = yield cmd_dict


def load_snapshot(
        tiles: typing.List[PictureTile]
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SnapshotId]:
    '''
    Returns the snapshot identifier.

    :param tiles: An array of tiles composing the snapshot.
    :returns: The id of the snapshot.
    '''
    params: T_JSON_DICT = dict()
    params['tiles'] = [i.to_json() for i in tiles]
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.loadSnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return SnapshotId.from_json(json['snapshotId'])


def make_snapshot(
        layer_id: LayerId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SnapshotId]:
    '''
    Returns the layer snapshot identifier.

    :param layer_id: The id of the layer.
    :returns: The id of the layer snapshot.
    '''
    params: T_JSON_DICT = dict()
    params['layerId'] = layer_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.makeSnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return SnapshotId.from_json(json['snapshotId'])


def profile_snapshot(
        snapshot_id: SnapshotId,
        min_repeat_count: typing.Optional[int] = None,
        min_duration: typing.Optional[float] = None,
        clip_rect: typing.Optional[dom.Rect] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[PaintProfile]]:
    '''
    :param snapshot_id: The id of the layer snapshot.
    :param min_repeat_count: *(Optional)* The maximum number of times to replay the snapshot (1, if not specified).
    :param min_duration: *(Optional)* The minimum duration (in seconds) to replay the snapshot.
    :param clip_rect: *(Optional)* The clip rectangle to apply when replaying the snapshot.
    :returns: The array of paint profiles, one per run.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    if min_repeat_count is not None:
        params['minRepeatCount'] = min_repeat_count
    if min_duration is not None:
        params['minDuration'] = min_duration
    if clip_rect is not None:
        params['clipRect'] = clip_rect.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.profileSnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return [PaintProfile.from_json(i) for i in json['timings']]


def release_snapshot(
        snapshot_id: SnapshotId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Releases layer snapshot captured by the back-end.

    :param snapshot_id: The id of the layer snapshot.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.releaseSnapshot',
        'params': params,
    }
    json = yield cmd_dict


def replay_snapshot(
        snapshot_id: SnapshotId,
        from_step: typing.Optional[int] = None,
        to_step: typing.Optional[int] = None,
        scale: typing.Optional[float] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,str]:
    '''
    Replays the layer snapshot and returns the resulting bitmap.

    :param snapshot_id: The id of the layer snapshot.
    :param from_step: *(Optional)* The first step to replay from (replay from the very start if not specified).
    :param to_step: *(Optional)* The last step to replay to (replay till the end if not specified).
    :param scale: *(Optional)* The scale to apply while replaying (defaults to 1).
    :returns: A data: URL for resulting image.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    if from_step is not None:
        params['fromStep'] = from_step
    if to_step is not None:
        params['toStep'] = to_step
    if scale is not None:
        params['scale'] = scale
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.replaySnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return str(json['dataURL'])


def snapshot_command_log(
        snapshot_id: SnapshotId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[dict]]:
    '''
    Replays the layer snapshot and returns canvas log.

    :param snapshot_id: The id of the layer snapshot.
    :returns: The array of canvas function calls.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.snapshotCommandLog',
        'params': params,
    }
    json = yield cmd_dict
    return [dict(i) for i in json['commandLog']]


@event_class('LayerTree.layerPainted')
@dataclass
class LayerPainted:
    #: The id of the painted layer.
    layer_id: LayerId
    #: Clip rectangle.
    clip: dom.Rect

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> LayerPainted:
        return cls(
            layer_id=LayerId.from_json(json['layerId']),
            clip=dom.Rect.from_json(json['clip'])
        )


@event_class('LayerTree.layerTreeDidChange')
@dataclass
class LayerTreeDidChange:
    #: Layer tree, absent if not in the compositing mode.
    layers: typing.Optional[typing.List[Layer]]

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> LayerTreeDidChange:
        return cls(
            layers=[Layer.from_json(i) for i in json['layers']] if 'layers' in json else None
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v137\layer_tree.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: LayerTree (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing
from . import dom


class LayerId(str):
    '''
    Unique Layer identifier.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> LayerId:
        return cls(json)

    def __repr__(self):
        return 'LayerId({})'.format(super().__repr__())


class SnapshotId(str):
    '''
    Unique snapshot identifier.
    '''
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> SnapshotId:
        return cls(json)

    def __repr__(self):
        return 'SnapshotId({})'.format(super().__repr__())


@dataclass
class ScrollRect:
    '''
    Rectangle where scrolling happens on the main thread.
    '''
    #: Rectangle itself.
    rect: dom.Rect

    #: Reason for rectangle to force scrolling on the main thread
    type_: str

    def to_json(self):
        json = dict()
        json['rect'] = self.rect.to_json()
        json['type'] = self.type_
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            rect=dom.Rect.from_json(json['rect']),
            type_=str(json['type']),
        )


@dataclass
class StickyPositionConstraint:
    '''
    Sticky position constraints.
    '''
    #: Layout rectangle of the sticky element before being shifted
    sticky_box_rect: dom.Rect

    #: Layout rectangle of the containing block of the sticky element
    containing_block_rect: dom.Rect

    #: The nearest sticky layer that shifts the sticky box
    nearest_layer_shifting_sticky_box: typing.Optional[LayerId] = None

    #: The nearest sticky layer that shifts the containing block
    nearest_layer_shifting_containing_block: typing.Optional[LayerId] = None

    def to_json(self):
        json = dict()
        json['stickyBoxRect'] = self.sticky_box_rect.to_json()
        json['containingBlockRect'] = self.containing_block_rect.to_json()
        if self.nearest_layer_shifting_sticky_box is not None:
            json['nearestLayerShiftingStickyBox'] = self.nearest_layer_shifting_sticky_box.to_json()
        if self.nearest_layer_shifting_containing_block is not None:
            json['nearestLayerShiftingContainingBlock'] = self.nearest_layer_shifting_containing_block.to_json()
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            sticky_box_rect=dom.Rect.from_json(json['stickyBoxRect']),
            containing_block_rect=dom.Rect.from_json(json['containingBlockRect']),
            nearest_layer_shifting_sticky_box=LayerId.from_json(json['nearestLayerShiftingStickyBox']) if 'nearestLayerShiftingStickyBox' in json else None,
            nearest_layer_shifting_containing_block=LayerId.from_json(json['nearestLayerShiftingContainingBlock']) if 'nearestLayerShiftingContainingBlock' in json else None,
        )


@dataclass
class PictureTile:
    '''
    Serialized fragment of layer picture along with its offset within the layer.
    '''
    #: Offset from owning layer left boundary
    x: float

    #: Offset from owning layer top boundary
    y: float

    #: Base64-encoded snapshot data.
    picture: str

    def to_json(self):
        json = dict()
        json['x'] = self.x
        json['y'] = self.y
        json['picture'] = self.picture
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            x=float(json['x']),
            y=float(json['y']),
            picture=str(json['picture']),
        )


@dataclass
class Layer:
    '''
    Information about a compositing layer.
    '''
    #: The unique id for this layer.
    layer_id: LayerId

    #: Offset from parent layer, X coordinate.
    offset_x: float

    #: Offset from parent layer, Y coordinate.
    offset_y: float

    #: Layer width.
    width: float

    #: Layer height.
    height: float

    #: Indicates how many time this layer has painted.
    paint_count: int

    #: Indicates whether this layer hosts any content, rather than being used for
    #: transform/scrolling purposes only.
    draws_content: bool

    #: The id of parent (not present for root).
    parent_layer_id: typing.Optional[LayerId] = None

    #: The backend id for the node associated with this layer.
    backend_node_id: typing.Optional[dom.BackendNodeId] = None

    #: Transformation matrix for layer, default is identity matrix
    transform: typing.Optional[typing.List[float]] = None

    #: Transform anchor point X, absent if no transform specified
    anchor_x: typing.Optional[float] = None

    #: Transform anchor point Y, absent if no transform specified
    anchor_y: typing.Optional[float] = None

    #: Transform anchor point Z, absent if no transform specified
    anchor_z: typing.Optional[float] = None

    #: Set if layer is not visible.
    invisible: typing.Optional[bool] = None

    #: Rectangles scrolling on main thread only.
    scroll_rects: typing.Optional[typing.List[ScrollRect]] = None

    #: Sticky position constraint information
    sticky_position_constraint: typing.Optional[StickyPositionConstraint] = None

    def to_json(self):
        json = dict()
        json['layerId'] = self.layer_id.to_json()
        json['offsetX'] = self.offset_x
        json['offsetY'] = self.offset_y
        json['width'] = self.width
        json['height'] = self.height
        json['paintCount'] = self.paint_count
        json['drawsContent'] = self.draws_content
        if self.parent_layer_id is not None:
            json['parentLayerId'] = self.parent_layer_id.to_json()
        if self.backend_node_id is not None:
            json['backendNodeId'] = self.backend_node_id.to_json()
        if self.transform is not None:
            json['transform'] = [i for i in self.transform]
        if self.anchor_x is not None:
            json['anchorX'] = self.anchor_x
        if self.anchor_y is not None:
            json['anchorY'] = self.anchor_y
        if self.anchor_z is not None:
            json['anchorZ'] = self.anchor_z
        if self.invisible is not None:
            json['invisible'] = self.invisible
        if self.scroll_rects is not None:
            json['scrollRects'] = [i.to_json() for i in self.scroll_rects]
        if self.sticky_position_constraint is not None:
            json['stickyPositionConstraint'] = self.sticky_position_constraint.to_json()
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            layer_id=LayerId.from_json(json['layerId']),
            offset_x=float(json['offsetX']),
            offset_y=float(json['offsetY']),
            width=float(json['width']),
            height=float(json['height']),
            paint_count=int(json['paintCount']),
            draws_content=bool(json['drawsContent']),
            parent_layer_id=LayerId.from_json(json['parentLayerId']) if 'parentLayerId' in json else None,
            backend_node_id=dom.BackendNodeId.from_json(json['backendNodeId']) if 'backendNodeId' in json else None,
            transform=[float(i) for i in json['transform']] if 'transform' in json else None,
            anchor_x=float(json['anchorX']) if 'anchorX' in json else None,
            anchor_y=float(json['anchorY']) if 'anchorY' in json else None,
            anchor_z=float(json['anchorZ']) if 'anchorZ' in json else None,
            invisible=bool(json['invisible']) if 'invisible' in json else None,
            scroll_rects=[ScrollRect.from_json(i) for i in json['scrollRects']] if 'scrollRects' in json else None,
            sticky_position_constraint=StickyPositionConstraint.from_json(json['stickyPositionConstraint']) if 'stickyPositionConstraint' in json else None,
        )


class PaintProfile(list):
    '''
    Array of timings, one per paint step.
    '''
    def to_json(self) -> typing.List[float]:
        return self

    @classmethod
    def from_json(cls, json: typing.List[float]) -> PaintProfile:
        return cls(json)

    def __repr__(self):
        return 'PaintProfile({})'.format(super().__repr__())


def compositing_reasons(
        layer_id: LayerId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.Tuple[typing.List[str], typing.List[str]]]:
    '''
    Provides the reasons why the given layer was composited.

    :param layer_id: The id of the layer for which we want to get the reasons it was composited.
    :returns: A tuple with the following items:

        0. **compositingReasons** - A list of strings specifying reasons for the given layer to become composited.
        1. **compositingReasonIds** - A list of strings specifying reason IDs for the given layer to become composited.
    '''
    params: T_JSON_DICT = dict()
    params['layerId'] = layer_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.compositingReasons',
        'params': params,
    }
    json = yield cmd_dict
    return (
        [str(i) for i in json['compositingReasons']],
        [str(i) for i in json['compositingReasonIds']]
    )


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disables compositing tree inspection.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.disable',
    }
    json = yield cmd_dict


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enables compositing tree inspection.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.enable',
    }
    json = yield cmd_dict


def load_snapshot(
        tiles: typing.List[PictureTile]
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SnapshotId]:
    '''
    Returns the snapshot identifier.

    :param tiles: An array of tiles composing the snapshot.
    :returns: The id of the snapshot.
    '''
    params: T_JSON_DICT = dict()
    params['tiles'] = [i.to_json() for i in tiles]
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.loadSnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return SnapshotId.from_json(json['snapshotId'])


def make_snapshot(
        layer_id: LayerId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SnapshotId]:
    '''
    Returns the layer snapshot identifier.

    :param layer_id: The id of the layer.
    :returns: The id of the layer snapshot.
    '''
    params: T_JSON_DICT = dict()
    params['layerId'] = layer_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.makeSnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return SnapshotId.from_json(json['snapshotId'])


def profile_snapshot(
        snapshot_id: SnapshotId,
        min_repeat_count: typing.Optional[int] = None,
        min_duration: typing.Optional[float] = None,
        clip_rect: typing.Optional[dom.Rect] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[PaintProfile]]:
    '''
    :param snapshot_id: The id of the layer snapshot.
    :param min_repeat_count: *(Optional)* The maximum number of times to replay the snapshot (1, if not specified).
    :param min_duration: *(Optional)* The minimum duration (in seconds) to replay the snapshot.
    :param clip_rect: *(Optional)* The clip rectangle to apply when replaying the snapshot.
    :returns: The array of paint profiles, one per run.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    if min_repeat_count is not None:
        params['minRepeatCount'] = min_repeat_count
    if min_duration is not None:
        params['minDuration'] = min_duration
    if clip_rect is not None:
        params['clipRect'] = clip_rect.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.profileSnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return [PaintProfile.from_json(i) for i in json['timings']]


def release_snapshot(
        snapshot_id: SnapshotId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Releases layer snapshot captured by the back-end.

    :param snapshot_id: The id of the layer snapshot.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.releaseSnapshot',
        'params': params,
    }
    json = yield cmd_dict


def replay_snapshot(
        snapshot_id: SnapshotId,
        from_step: typing.Optional[int] = None,
        to_step: typing.Optional[int] = None,
        scale: typing.Optional[float] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,str]:
    '''
    Replays the layer snapshot and returns the resulting bitmap.

    :param snapshot_id: The id of the layer snapshot.
    :param from_step: *(Optional)* The first step to replay from (replay from the very start if not specified).
    :param to_step: *(Optional)* The last step to replay to (replay till the end if not specified).
    :param scale: *(Optional)* The scale to apply while replaying (defaults to 1).
    :returns: A data: URL for resulting image.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    if from_step is not None:
        params['fromStep'] = from_step
    if to_step is not None:
        params['toStep'] = to_step
    if scale is not None:
        params['scale'] = scale
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.replaySnapshot',
        'params': params,
    }
    json = yield cmd_dict
    return str(json['dataURL'])


def snapshot_command_log(
        snapshot_id: SnapshotId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[dict]]:
    '''
    Replays the layer snapshot and returns canvas log.

    :param snapshot_id: The id of the layer snapshot.
    :returns: The array of canvas function calls.
    '''
    params: T_JSON_DICT = dict()
    params['snapshotId'] = snapshot_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'LayerTree.snapshotCommandLog',
        'params': params,
    }
    json = yield cmd_dict
    return [dict(i) for i in json['commandLog']]


@event_class('LayerTree.layerPainted')
@dataclass
class LayerPainted:
    #: The id of the painted layer.
    layer_id: LayerId
    #: Clip rectangle.
    clip: dom.Rect

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> LayerPainted:
        return cls(
            layer_id=LayerId.from_json(json['layerId']),
            clip=dom.Rect.from_json(json['clip'])
        )


@event_class('LayerTree.layerTreeDidChange')
@dataclass
class LayerTreeDidChange:
    #: Layer tree, absent if not in the compositing mode.
    layers: typing.Optional[typing.List[Layer]]

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> LayerTreeDidChange:
        return cls(
            layers=[Layer.from_json(i) for i in json['layers']] if 'layers' in json else None
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\interchange\column.py ===
from __future__ import annotations

from typing import (
    TYPE_CHECKING,
    Any,
)

import numpy as np

from pandas._libs.lib import infer_dtype
from pandas._libs.tslibs import iNaT
from pandas.errors import NoBufferPresent
from pandas.util._decorators import cache_readonly

from pandas.core.dtypes.dtypes import BaseMaskedDtype

import pandas as pd
from pandas import (
    ArrowDtype,
    DatetimeTZDtype,
)
from pandas.api.types import is_string_dtype
from pandas.core.interchange.buffer import (
    PandasBuffer,
    PandasBufferPyarrow,
)
from pandas.core.interchange.dataframe_protocol import (
    Column,
    ColumnBuffers,
    ColumnNullType,
    DtypeKind,
)
from pandas.core.interchange.utils import (
    ArrowCTypes,
    Endianness,
    dtype_to_arrow_c_fmt,
)

if TYPE_CHECKING:
    from pandas.core.interchange.dataframe_protocol import Buffer

_NP_KINDS = {
    "i": DtypeKind.INT,
    "u": DtypeKind.UINT,
    "f": DtypeKind.FLOAT,
    "b": DtypeKind.BOOL,
    "U": DtypeKind.STRING,
    "M": DtypeKind.DATETIME,
    "m": DtypeKind.DATETIME,
}

_NULL_DESCRIPTION = {
    DtypeKind.FLOAT: (ColumnNullType.USE_NAN, None),
    DtypeKind.DATETIME: (ColumnNullType.USE_SENTINEL, iNaT),
    DtypeKind.INT: (ColumnNullType.NON_NULLABLE, None),
    DtypeKind.UINT: (ColumnNullType.NON_NULLABLE, None),
    DtypeKind.BOOL: (ColumnNullType.NON_NULLABLE, None),
    # Null values for categoricals are stored as `-1` sentinel values
    # in the category date (e.g., `col.values.codes` is int8 np.ndarray)
    DtypeKind.CATEGORICAL: (ColumnNullType.USE_SENTINEL, -1),
    # follow Arrow in using 1 as valid value and 0 for missing/null value
    DtypeKind.STRING: (ColumnNullType.USE_BYTEMASK, 0),
}

_NO_VALIDITY_BUFFER = {
    ColumnNullType.NON_NULLABLE: "This column is non-nullable",
    ColumnNullType.USE_NAN: "This column uses NaN as null",
    ColumnNullType.USE_SENTINEL: "This column uses a sentinel value",
}


class PandasColumn(Column):
    """
    A column object, with only the methods and properties required by the
    interchange protocol defined.
    A column can contain one or more chunks. Each chunk can contain up to three
    buffers - a data buffer, a mask buffer (depending on null representation),
    and an offsets buffer (if variable-size binary; e.g., variable-length
    strings).
    Note: this Column object can only be produced by ``__dataframe__``, so
          doesn't need its own version or ``__column__`` protocol.
    """

    def __init__(self, column: pd.Series, allow_copy: bool = True) -> None:
        """
        Note: doesn't deal with extension arrays yet, just assume a regular
        Series/ndarray for now.
        """
        if isinstance(column, pd.DataFrame):
            raise TypeError(
                "Expected a Series, got a DataFrame. This likely happened "
                "because you called __dataframe__ on a DataFrame which, "
                "after converting column names to string, resulted in duplicated "
                f"names: {column.columns}. Please rename these columns before "
                "using the interchange protocol."
            )
        if not isinstance(column, pd.Series):
            raise NotImplementedError(f"Columns of type {type(column)} not handled yet")

        # Store the column as a private attribute
        self._col = column
        self._allow_copy = allow_copy

    def size(self) -> int:
        """
        Size of the column, in elements.
        """
        return self._col.size

    @property
    def offset(self) -> int:
        """
        Offset of first element. Always zero.
        """
        # TODO: chunks are implemented now, probably this should return something
        return 0

    @cache_readonly
    def dtype(self) -> tuple[DtypeKind, int, str, str]:
        dtype = self._col.dtype

        if isinstance(dtype, pd.CategoricalDtype):
            codes = self._col.values.codes
            (
                _,
                bitwidth,
                c_arrow_dtype_f_str,
                _,
            ) = self._dtype_from_pandasdtype(codes.dtype)
            return (
                DtypeKind.CATEGORICAL,
                bitwidth,
                c_arrow_dtype_f_str,
                Endianness.NATIVE,
            )
        elif is_string_dtype(dtype):
            if infer_dtype(self._col) in ("string", "empty"):
                return (
                    DtypeKind.STRING,
                    8,
                    dtype_to_arrow_c_fmt(dtype),
                    Endianness.NATIVE,
                )
            raise NotImplementedError("Non-string object dtypes are not supported yet")
        else:
            return self._dtype_from_pandasdtype(dtype)

    def _dtype_from_pandasdtype(self, dtype) -> tuple[DtypeKind, int, str, str]:
        """
        See `self.dtype` for details.
        """
        # Note: 'c' (complex) not handled yet (not in array spec v1).
        #       'b', 'B' (bytes), 'S', 'a', (old-style string) 'V' (void) not handled
        #       datetime and timedelta both map to datetime (is timedelta handled?)

        kind = _NP_KINDS.get(dtype.kind, None)
        if kind is None:
            # Not a NumPy dtype. Check if it's a categorical maybe
            raise ValueError(f"Data type {dtype} not supported by interchange protocol")
        if isinstance(dtype, ArrowDtype):
            byteorder = dtype.numpy_dtype.byteorder
        elif isinstance(dtype, DatetimeTZDtype):
            byteorder = dtype.base.byteorder  # type: ignore[union-attr]
        elif isinstance(dtype, BaseMaskedDtype):
            byteorder = dtype.numpy_dtype.byteorder
        else:
            byteorder = dtype.byteorder

        if dtype == "bool[pyarrow]":
            # return early to avoid the `* 8` below, as this is a bitmask
            # rather than a bytemask
            return (
                kind,
                dtype.itemsize,  # pyright: ignore[reportGeneralTypeIssues]
                ArrowCTypes.BOOL,
                byteorder,
            )

        return kind, dtype.itemsize * 8, dtype_to_arrow_c_fmt(dtype), byteorder

    @property
    def describe_categorical(self):
        """
        If the dtype is categorical, there are two options:
        - There are only values in the data buffer.
        - There is a separate non-categorical Column encoding for categorical values.

        Raises TypeError if the dtype is not categorical

        Content of returned dict:
            - "is_ordered" : bool, whether the ordering of dictionary indices is
                             semantically meaningful.
            - "is_dictionary" : bool, whether a dictionary-style mapping of
                                categorical values to other objects exists
            - "categories" : Column representing the (implicit) mapping of indices to
                             category values (e.g. an array of cat1, cat2, ...).
                             None if not a dictionary-style categorical.
        """
        if not self.dtype[0] == DtypeKind.CATEGORICAL:
            raise TypeError(
                "describe_categorical only works on a column with categorical dtype!"
            )

        return {
            "is_ordered": self._col.cat.ordered,
            "is_dictionary": True,
            "categories": PandasColumn(pd.Series(self._col.cat.categories)),
        }

    @property
    def describe_null(self):
        if isinstance(self._col.dtype, BaseMaskedDtype):
            column_null_dtype = ColumnNullType.USE_BYTEMASK
            null_value = 1
            return column_null_dtype, null_value
        if isinstance(self._col.dtype, ArrowDtype):
            # We already rechunk (if necessary / allowed) upon initialization, so this
            # is already single-chunk by the time we get here.
            if self._col.array._pa_array.chunks[0].buffers()[0] is None:  # type: ignore[attr-defined]
                return ColumnNullType.NON_NULLABLE, None
            return ColumnNullType.USE_BITMASK, 0
        kind = self.dtype[0]
        try:
            null, value = _NULL_DESCRIPTION[kind]
        except KeyError:
            raise NotImplementedError(f"Data type {kind} not yet supported")

        return null, value

    @cache_readonly
    def null_count(self) -> int:
        """
        Number of null elements. Should always be known.
        """
        return self._col.isna().sum().item()

    @property
    def metadata(self) -> dict[str, pd.Index]:
        """
        Store specific metadata of the column.
        """
        return {"pandas.index": self._col.index}

    def num_chunks(self) -> int:
        """
        Return the number of chunks the column consists of.
        """
        return 1

    def get_chunks(self, n_chunks: int | None = None):
        """
        Return an iterator yielding the chunks.
        See `DataFrame.get_chunks` for details on ``n_chunks``.
        """
        if n_chunks and n_chunks > 1:
            size = len(self._col)
            step = size // n_chunks
            if size % n_chunks != 0:
                step += 1
            for start in range(0, step * n_chunks, step):
                yield PandasColumn(
                    self._col.iloc[start : start + step], self._allow_copy
                )
        else:
            yield self

    def get_buffers(self) -> ColumnBuffers:
        """
        Return a dictionary containing the underlying buffers.
        The returned dictionary has the following contents:
            - "data": a two-element tuple whose first element is a buffer
                      containing the data and whose second element is the data
                      buffer's associated dtype.
            - "validity": a two-element tuple whose first element is a buffer
                          containing mask values indicating missing data and
                          whose second element is the mask value buffer's
                          associated dtype. None if the null representation is
                          not a bit or byte mask.
            - "offsets": a two-element tuple whose first element is a buffer
                         containing the offset values for variable-size binary
                         data (e.g., variable-length strings) and whose second
                         element is the offsets buffer's associated dtype. None
                         if the data buffer does not have an associated offsets
                         buffer.
        """
        buffers: ColumnBuffers = {
            "data": self._get_data_buffer(),
            "validity": None,
            "offsets": None,
        }

        try:
            buffers["validity"] = self._get_validity_buffer()
        except NoBufferPresent:
            pass

        try:
            buffers["offsets"] = self._get_offsets_buffer()
        except NoBufferPresent:
            pass

        return buffers

    def _get_data_buffer(
        self,
    ) -> tuple[Buffer, tuple[DtypeKind, int, str, str]]:
        """
        Return the buffer containing the data and the buffer's associated dtype.
        """
        buffer: Buffer
        if self.dtype[0] in (
            DtypeKind.INT,
            DtypeKind.UINT,
            DtypeKind.FLOAT,
            DtypeKind.BOOL,
            DtypeKind.DATETIME,
        ):
            # self.dtype[2] is an ArrowCTypes.TIMESTAMP where the tz will make
            # it longer than 4 characters
            dtype = self.dtype
            if self.dtype[0] == DtypeKind.DATETIME and len(self.dtype[2]) > 4:
                np_arr = self._col.dt.tz_convert(None).to_numpy()
            else:
                arr = self._col.array
                if isinstance(self._col.dtype, BaseMaskedDtype):
                    np_arr = arr._data  # type: ignore[attr-defined]
                elif isinstance(self._col.dtype, ArrowDtype):
                    # We already rechunk (if necessary / allowed) upon initialization,
                    # so this is already single-chunk by the time we get here.
                    arr = arr._pa_array.chunks[0]  # type: ignore[attr-defined]
                    buffer = PandasBufferPyarrow(
                        arr.buffers()[1],  # type: ignore[attr-defined]
                        length=len(arr),
                    )
                    return buffer, dtype
                else:
                    np_arr = arr._ndarray  # type: ignore[attr-defined]
            buffer = PandasBuffer(np_arr, allow_copy=self._allow_copy)
        elif self.dtype[0] == DtypeKind.CATEGORICAL:
            codes = self._col.values._codes
            buffer = PandasBuffer(codes, allow_copy=self._allow_copy)
            dtype = self._dtype_from_pandasdtype(codes.dtype)
        elif self.dtype[0] == DtypeKind.STRING:
            # Marshal the strings from a NumPy object array into a byte array
            buf = self._col.to_numpy()
            b = bytearray()

            # TODO: this for-loop is slow; can be implemented in Cython/C/C++ later
            for obj in buf:
                if isinstance(obj, str):
                    b.extend(obj.encode(encoding="utf-8"))

            # Convert the byte array to a Pandas "buffer" using
            # a NumPy array as the backing store
            buffer = PandasBuffer(np.frombuffer(b, dtype="uint8"))

            # Define the dtype for the returned buffer
            # TODO: this will need correcting
            # https://github.com/pandas-dev/pandas/issues/54781
            dtype = self.dtype
        else:
            raise NotImplementedError(f"Data type {self._col.dtype} not handled yet")

        return buffer, dtype

    def _get_validity_buffer(self) -> tuple[Buffer, Any] | None:
        """
        Return the buffer containing the mask values indicating missing data and
        the buffer's associated dtype.
        Raises NoBufferPresent if null representation is not a bit or byte mask.
        """
        null, invalid = self.describe_null
        buffer: Buffer
        if isinstance(self._col.dtype, ArrowDtype):
            # We already rechunk (if necessary / allowed) upon initialization, so this
            # is already single-chunk by the time we get here.
            arr = self._col.array._pa_array.chunks[0]  # type: ignore[attr-defined]
            dtype = (DtypeKind.BOOL, 1, ArrowCTypes.BOOL, Endianness.NATIVE)
            if arr.buffers()[0] is None:
                return None
            buffer = PandasBufferPyarrow(
                arr.buffers()[0],
                length=len(arr),
            )
            return buffer, dtype

        if isinstance(self._col.dtype, BaseMaskedDtype):
            mask = self._col.array._mask  # type: ignore[attr-defined]
            buffer = PandasBuffer(mask)
            dtype = (DtypeKind.BOOL, 8, ArrowCTypes.BOOL, Endianness.NATIVE)
            return buffer, dtype

        if self.dtype[0] == DtypeKind.STRING:
            # For now, use byte array as the mask.
            # TODO: maybe store as bit array to save space?..
            buf = self._col.to_numpy()

            # Determine the encoding for valid values
            valid = invalid == 0
            invalid = not valid

            mask = np.zeros(shape=(len(buf),), dtype=np.bool_)
            for i, obj in enumerate(buf):
                mask[i] = valid if isinstance(obj, str) else invalid

            # Convert the mask array to a Pandas "buffer" using
            # a NumPy array as the backing store
            buffer = PandasBuffer(mask)

            # Define the dtype of the returned buffer
            dtype = (DtypeKind.BOOL, 8, ArrowCTypes.BOOL, Endianness.NATIVE)

            return buffer, dtype

        try:
            msg = f"{_NO_VALIDITY_BUFFER[null]} so does not have a separate mask"
        except KeyError:
            # TODO: implement for other bit/byte masks?
            raise NotImplementedError("See self.describe_null")

        raise NoBufferPresent(msg)

    def _get_offsets_buffer(self) -> tuple[PandasBuffer, Any]:
        """
        Return the buffer containing the offset values for variable-size binary
        data (e.g., variable-length strings) and the buffer's associated dtype.
        Raises NoBufferPresent if the data buffer does not have an associated
        offsets buffer.
        """
        if self.dtype[0] == DtypeKind.STRING:
            # For each string, we need to manually determine the next offset
            values = self._col.to_numpy()
            ptr = 0
            offsets = np.zeros(shape=(len(values) + 1,), dtype=np.int64)
            for i, v in enumerate(values):
                # For missing values (in this case, `np.nan` values)
                # we don't increment the pointer
                if isinstance(v, str):
                    b = v.encode(encoding="utf-8")
                    ptr += len(b)

                offsets[i + 1] = ptr

            # Convert the offsets to a Pandas "buffer" using
            # the NumPy array as the backing store
            buffer = PandasBuffer(offsets)

            # Assemble the buffer dtype info
            dtype = (
                DtypeKind.INT,
                64,
                ArrowCTypes.INT64,
                Endianness.NATIVE,
            )  # note: currently only support native endianness
        else:
            raise NoBufferPresent(
                "This column has a fixed-length dtype so "
                "it does not have an offsets buffer"
            )

        return buffer, dtype

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\testing\runtests_pytest.py ===
"""Backwards compatible functions for running tests from SymPy using pytest.

SymPy historically had its own testing framework that aimed to:
- be compatible with pytest;
- operate similarly (or identically) to pytest;
- not require any external dependencies;
- have all the functionality in one file only;
- have no magic, just import the test file and execute the test functions; and
- be portable.

To reduce the maintenance burden of developing an independent testing framework
and to leverage the benefits of existing Python testing infrastructure, SymPy
now uses pytest (and various of its plugins) to run the test suite.

To maintain backwards compatibility with the legacy testing interface of SymPy,
which implemented functions that allowed users to run the tests on their
installed version of SymPy, the functions in this module are implemented to
match the existing API while thinly wrapping pytest.

These two key functions are `test` and `doctest`.

"""

import functools
import importlib.util
import os
import pathlib
import re
from fnmatch import fnmatch
from typing import List, Optional, Tuple

try:
    import pytest
except ImportError:

    class NoPytestError(Exception):
        """Raise when an internal test helper function is called with pytest."""

    class pytest:  # type: ignore
        """Shadow to support pytest features when pytest can't be imported."""

        @staticmethod
        def main(*args, **kwargs):
            msg = 'pytest must be installed to run tests via this function'
            raise NoPytestError(msg)

from sympy.testing.runtests import test as test_sympy


TESTPATHS_DEFAULT = (
    pathlib.Path('sympy'),
    pathlib.Path('doc', 'src'),
)
BLACKLIST_DEFAULT = (
    'sympy/integrals/rubi/rubi_tests/tests',
)


class PytestPluginManager:
    """Module names for pytest plugins used by SymPy."""
    PYTEST: str = 'pytest'
    RANDOMLY: str = 'pytest_randomly'
    SPLIT: str = 'pytest_split'
    TIMEOUT: str = 'pytest_timeout'
    XDIST: str = 'xdist'

    @functools.cached_property
    def has_pytest(self) -> bool:
        return bool(importlib.util.find_spec(self.PYTEST))

    @functools.cached_property
    def has_randomly(self) -> bool:
        return bool(importlib.util.find_spec(self.RANDOMLY))

    @functools.cached_property
    def has_split(self) -> bool:
        return bool(importlib.util.find_spec(self.SPLIT))

    @functools.cached_property
    def has_timeout(self) -> bool:
        return bool(importlib.util.find_spec(self.TIMEOUT))

    @functools.cached_property
    def has_xdist(self) -> bool:
        return bool(importlib.util.find_spec(self.XDIST))


split_pattern = re.compile(r'([1-9][0-9]*)/([1-9][0-9]*)')


@functools.lru_cache
def sympy_dir() -> pathlib.Path:
    """Returns the root SymPy directory."""
    return pathlib.Path(__file__).parents[2]


def update_args_with_paths(
    paths: List[str],
    keywords: Optional[Tuple[str]],
    args: List[str],
) -> List[str]:
    """Appends valid paths and flags to the args `list` passed to `pytest.main`.

    The are three different types of "path" that a user may pass to the `paths`
    positional arguments, all of which need to be handled slightly differently:

    1. Nothing is passed
        The paths to the `testpaths` defined in `pytest.ini` need to be appended
        to the arguments list.
    2. Full, valid paths are passed
        These paths need to be validated but can then be directly appended to
        the arguments list.
    3. Partial paths are passed.
        The `testpaths` defined in `pytest.ini` need to be recursed and any
        matches be appended to the arguments list.

    """

    def find_paths_matching_partial(partial_paths):
        partial_path_file_patterns = []
        for partial_path in partial_paths:
            if len(partial_path) >= 4:
                has_test_prefix = partial_path[:4] == 'test'
                has_py_suffix = partial_path[-3:] == '.py'
            elif len(partial_path) >= 3:
                has_test_prefix = False
                has_py_suffix = partial_path[-3:] == '.py'
            else:
                has_test_prefix = False
                has_py_suffix = False
            if has_test_prefix and has_py_suffix:
                partial_path_file_patterns.append(partial_path)
            elif has_test_prefix:
                partial_path_file_patterns.append(f'{partial_path}*.py')
            elif has_py_suffix:
                partial_path_file_patterns.append(f'test*{partial_path}')
            else:
                partial_path_file_patterns.append(f'test*{partial_path}*.py')
        matches = []
        for testpath in valid_testpaths_default:
            for path, dirs, files in os.walk(testpath, topdown=True):
                zipped = zip(partial_paths, partial_path_file_patterns)
                for (partial_path, partial_path_file) in zipped:
                    if fnmatch(path, f'*{partial_path}*'):
                        matches.append(str(pathlib.Path(path)))
                        dirs[:] = []
                    else:
                        for file in files:
                            if fnmatch(file, partial_path_file):
                                matches.append(str(pathlib.Path(path, file)))
        return matches

    def is_tests_file(filepath: str) -> bool:
        path = pathlib.Path(filepath)
        if not path.is_file():
            return False
        if not path.parts[-1].startswith('test_'):
            return False
        if not path.suffix == '.py':
            return False
        return True

    def find_tests_matching_keywords(keywords, filepath):
        matches = []
        source = pathlib.Path(filepath).read_text(encoding='utf-8')
        for line in source.splitlines():
            if line.lstrip().startswith('def '):
                for kw in keywords:
                    if line.lower().find(kw.lower()) != -1:
                        test_name = line.split(' ')[1].split('(')[0]
                        full_test_path = filepath + '::' + test_name
                        matches.append(full_test_path)
        return matches

    valid_testpaths_default = []
    for testpath in TESTPATHS_DEFAULT:
        absolute_testpath = pathlib.Path(sympy_dir(), testpath)
        if absolute_testpath.exists():
            valid_testpaths_default.append(str(absolute_testpath))

    candidate_paths = []
    if paths:
        full_paths = []
        partial_paths = []
        for path in paths:
            if pathlib.Path(path).exists():
                full_paths.append(str(pathlib.Path(sympy_dir(), path)))
            else:
                partial_paths.append(path)
        matched_paths = find_paths_matching_partial(partial_paths)
        candidate_paths.extend(full_paths)
        candidate_paths.extend(matched_paths)
    else:
        candidate_paths.extend(valid_testpaths_default)

    if keywords is not None and keywords != ():
        matches = []
        for path in candidate_paths:
            if is_tests_file(path):
                test_matches = find_tests_matching_keywords(keywords, path)
                matches.extend(test_matches)
            else:
                for root, dirnames, filenames in os.walk(path):
                    for filename in filenames:
                        absolute_filepath = str(pathlib.Path(root, filename))
                        if is_tests_file(absolute_filepath):
                            test_matches = find_tests_matching_keywords(
                                keywords,
                                absolute_filepath,
                            )
                            matches.extend(test_matches)
        args.extend(matches)
    else:
        args.extend(candidate_paths)

    return args


def make_absolute_path(partial_path: str) -> str:
    """Convert a partial path to an absolute path.

    A path such a `sympy/core` might be needed. However, absolute paths should
    be used in the arguments to pytest in all cases as it avoids errors that
    arise from nonexistent paths.

    This function assumes that partial_paths will be passed in such that they
    begin with the explicit `sympy` directory, i.e. `sympy/...`.

    """

    def is_valid_partial_path(partial_path: str) -> bool:
        """Assumption that partial paths are defined from the `sympy` root."""
        return pathlib.Path(partial_path).parts[0] == 'sympy'

    if not is_valid_partial_path(partial_path):
        msg = (
            f'Partial path {dir(partial_path)} is invalid, partial paths are '
            f'expected to be defined with the `sympy` directory as the root.'
        )
        raise ValueError(msg)

    absolute_path = str(pathlib.Path(sympy_dir(), partial_path))
    return absolute_path


def test(*paths, subprocess=True, rerun=0, **kwargs):
    """Interface to run tests via pytest compatible with SymPy's test runner.

    Explanation
    ===========

    Note that a `pytest.ExitCode`, which is an `enum`, is returned. This is
    different to the legacy SymPy test runner which would return a `bool`. If
    all tests successfully pass the `pytest.ExitCode.OK` with value `0` is
    returned, whereas the legacy SymPy test runner would return `True`. In any
    other scenario, a non-zero `enum` value is returned, whereas the legacy
    SymPy test runner would return `False`. Users need to, therefore, be careful
    if treating the pytest exit codes as booleans because
    `bool(pytest.ExitCode.OK)` evaluates to `False`, the opposite of legacy
    behaviour.

    Examples
    ========

    >>> import sympy  # doctest: +SKIP

    Run one file:

    >>> sympy.test('sympy/core/tests/test_basic.py')  # doctest: +SKIP
    >>> sympy.test('_basic')  # doctest: +SKIP

    Run all tests in sympy/functions/ and some particular file:

    >>> sympy.test("sympy/core/tests/test_basic.py",
    ...            "sympy/functions")  # doctest: +SKIP

    Run all tests in sympy/core and sympy/utilities:

    >>> sympy.test("/core", "/util")  # doctest: +SKIP

    Run specific test from a file:

    >>> sympy.test("sympy/core/tests/test_basic.py",
    ...            kw="test_equality")  # doctest: +SKIP

    Run specific test from any file:

    >>> sympy.test(kw="subs")  # doctest: +SKIP

    Run the tests using the legacy SymPy runner:

    >>> sympy.test(use_sympy_runner=True)  # doctest: +SKIP

    Note that this option is slated for deprecation in the near future and is
    only currently provided to ensure users have an alternative option while the
    pytest-based runner receives real-world testing.

    Parameters
    ==========
    paths : first n positional arguments of strings
        Paths, both partial and absolute, describing which subset(s) of the test
        suite are to be run.
    subprocess : bool, default is True
        Legacy option, is currently ignored.
    rerun : int, default is 0
        Legacy option, is ignored.
    use_sympy_runner : bool or None, default is None
        Temporary option to invoke the legacy SymPy test runner instead of
        `pytest.main`. Will be removed in the near future.
    verbose : bool, default is False
        Sets the verbosity of the pytest output. Using `True` will add the
        `--verbose` option to the pytest call.
    tb : str, 'auto', 'long', 'short', 'line', 'native', or 'no'
        Sets the traceback print mode of pytest using the `--tb` option.
    kw : str
        Only run tests which match the given substring expression. An expression
        is a Python evaluatable expression where all names are substring-matched
        against test names and their parent classes. Example: -k 'test_method or
        test_other' matches all test functions and classes whose name contains
        'test_method' or 'test_other', while -k 'not test_method' matches those
        that don't contain 'test_method' in their names. -k 'not test_method and
        not test_other' will eliminate the matches. Additionally keywords are
        matched to classes and functions containing extra names in their
        'extra_keyword_matches' set, as well as functions which have names
        assigned directly to them. The matching is case-insensitive.
    pdb : bool, default is False
        Start the interactive Python debugger on errors or `KeyboardInterrupt`.
    colors : bool, default is True
        Color terminal output.
    force_colors : bool, default is False
        Legacy option, is ignored.
    sort : bool, default is True
        Run the tests in sorted order. pytest uses a sorted test order by
        default. Requires pytest-randomly.
    seed : int
        Seed to use for random number generation. Requires pytest-randomly.
    timeout : int, default is 0
        Timeout in seconds before dumping the stacks. 0 means no timeout.
        Requires pytest-timeout.
    fail_on_timeout : bool, default is False
        Legacy option, is currently ignored.
    slow : bool, default is False
        Run the subset of tests marked as `slow`.
    enhance_asserts : bool, default is False
        Legacy option, is currently ignored.
    split : string in form `<SPLIT>/<GROUPS>` or None, default is None
        Used to split the tests up. As an example, if `split='2/3' is used then
        only the middle third of tests are run. Requires pytest-split.
    time_balance : bool, default is True
        Legacy option, is currently ignored.
    blacklist : iterable of test paths as strings, default is BLACKLIST_DEFAULT
        Blacklisted test paths are ignored using the `--ignore` option. Paths
        may be partial or absolute. If partial then they are matched against
        all paths in the pytest tests path.
    parallel : bool, default is False
        Parallelize the test running using pytest-xdist. If `True` then pytest
        will automatically detect the number of CPU cores available and use them
        all. Requires pytest-xdist.
    store_durations : bool, False
        Store test durations into the file `.test_durations`. The is used by
        `pytest-split` to help determine more even splits when more than one
        test group is being used. Requires pytest-split.

    """
    # NOTE: to be removed alongside SymPy test runner
    if kwargs.get('use_sympy_runner', False):
        kwargs.pop('parallel', False)
        kwargs.pop('store_durations', False)
        kwargs.pop('use_sympy_runner', True)
        if kwargs.get('slow') is None:
            kwargs['slow'] = False
        return test_sympy(*paths, subprocess=True, rerun=0, **kwargs)

    pytest_plugin_manager = PytestPluginManager()
    if not pytest_plugin_manager.has_pytest:
        pytest.main()

    args = []

    if kwargs.get('verbose', False):
        args.append('--verbose')

    if tb := kwargs.get('tb'):
        args.extend(['--tb', tb])

    if kwargs.get('pdb'):
        args.append('--pdb')

    if not kwargs.get('colors', True):
        args.extend(['--color', 'no'])

    if seed := kwargs.get('seed'):
        if not pytest_plugin_manager.has_randomly:
            msg = '`pytest-randomly` plugin required to control random seed.'
            raise ModuleNotFoundError(msg)
        args.extend(['--randomly-seed', str(seed)])

    if kwargs.get('sort', True) and pytest_plugin_manager.has_randomly:
        args.append('--randomly-dont-reorganize')
    elif not kwargs.get('sort', True) and not pytest_plugin_manager.has_randomly:
        msg = '`pytest-randomly` plugin required to randomize test order.'
        raise ModuleNotFoundError(msg)

    if timeout := kwargs.get('timeout', None):
        if not pytest_plugin_manager.has_timeout:
            msg = '`pytest-timeout` plugin required to apply timeout to tests.'
            raise ModuleNotFoundError(msg)
        args.extend(['--timeout', str(int(timeout))])

    # Skip slow tests by default and always skip tooslow tests
    if kwargs.get('slow', False):
        args.extend(['-m', 'slow and not tooslow'])
    else:
        args.extend(['-m', 'not slow and not tooslow'])

    if (split := kwargs.get('split')) is not None:
        if not pytest_plugin_manager.has_split:
            msg = '`pytest-split` plugin required to run tests as groups.'
            raise ModuleNotFoundError(msg)
        match = split_pattern.match(split)
        if not match:
            msg = ('split must be a string of the form a/b where a and b are '
                   'positive nonzero ints')
            raise ValueError(msg)
        group, splits = map(str, match.groups())
        args.extend(['--group', group, '--splits', splits])
        if group > splits:
            msg = (f'cannot have a group number {group} with only {splits} '
                   'splits')
            raise ValueError(msg)

    if blacklist := kwargs.get('blacklist', BLACKLIST_DEFAULT):
        for path in blacklist:
            args.extend(['--ignore', make_absolute_path(path)])

    if kwargs.get('parallel', False):
        if not pytest_plugin_manager.has_xdist:
            msg = '`pytest-xdist` plugin required to run tests in parallel.'
            raise ModuleNotFoundError(msg)
        args.extend(['-n', 'auto'])

    if kwargs.get('store_durations', False):
        if not pytest_plugin_manager.has_split:
            msg = '`pytest-split` plugin required to store test durations.'
            raise ModuleNotFoundError(msg)
        args.append('--store-durations')

    if (keywords := kwargs.get('kw')) is not None:
        keywords = tuple(str(kw) for kw in keywords)
    else:
        keywords = ()

    args = update_args_with_paths(paths, keywords, args)
    exit_code = pytest.main(args)
    return exit_code


def doctest():
    """Interface to run doctests via pytest compatible with SymPy's test runner.
    """
    raise NotImplementedError