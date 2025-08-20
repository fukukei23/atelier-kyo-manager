
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\bert_test_data.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

# It is a tool to generate test data for a bert model.
# The test data can be used by onnxruntime_perf_test tool to evaluate the inference latency.

import argparse
import os
import random
from pathlib import Path

import numpy as np
from onnx import ModelProto, TensorProto, numpy_helper
from onnx_model import OnnxModel


def fake_input_ids_data(
    input_ids: TensorProto, batch_size: int, sequence_length: int, dictionary_size: int
) -> np.ndarray:
    """Create input tensor based on the graph input of input_ids

    Args:
        input_ids (TensorProto): graph input of the input_ids input tensor
        batch_size (int): batch size
        sequence_length (int): sequence length
        dictionary_size (int): vocabulary size of dictionary

    Returns:
        np.ndarray: the input tensor created
    """
    assert input_ids.type.tensor_type.elem_type in [
        TensorProto.FLOAT,
        TensorProto.INT32,
        TensorProto.INT64,
    ]

    data = np.random.randint(dictionary_size, size=(batch_size, sequence_length), dtype=np.int32)

    if input_ids.type.tensor_type.elem_type == TensorProto.FLOAT:
        data = np.float32(data)
    elif input_ids.type.tensor_type.elem_type == TensorProto.INT64:
        data = np.int64(data)

    return data


def fake_segment_ids_data(segment_ids: TensorProto, batch_size: int, sequence_length: int) -> np.ndarray:
    """Create input tensor based on the graph input of segment_ids

    Args:
        segment_ids (TensorProto): graph input of the token_type_ids input tensor
        batch_size (int): batch size
        sequence_length (int): sequence length

    Returns:
        np.ndarray: the input tensor created
    """
    assert segment_ids.type.tensor_type.elem_type in [
        TensorProto.FLOAT,
        TensorProto.INT32,
        TensorProto.INT64,
    ]

    data = np.zeros((batch_size, sequence_length), dtype=np.int32)

    if segment_ids.type.tensor_type.elem_type == TensorProto.FLOAT:
        data = np.float32(data)
    elif segment_ids.type.tensor_type.elem_type == TensorProto.INT64:
        data = np.int64(data)

    return data


def get_random_length(max_sequence_length: int, average_sequence_length: int):
    assert average_sequence_length >= 1 and average_sequence_length <= max_sequence_length

    # For uniform distribution, we find proper lower and upper bounds so that the average is in the middle.
    if 2 * average_sequence_length > max_sequence_length:
        return random.randint(2 * average_sequence_length - max_sequence_length, max_sequence_length)
    else:
        return random.randint(1, 2 * average_sequence_length - 1)


def fake_input_mask_data(
    input_mask: TensorProto,
    batch_size: int,
    sequence_length: int,
    average_sequence_length: int,
    random_sequence_length: bool,
    mask_type: int = 2,
) -> np.ndarray:
    """Create input tensor based on the graph input of segment_ids.

    Args:
        input_mask (TensorProto): graph input of the attention mask input tensor
        batch_size (int): batch size
        sequence_length (int): sequence length
        average_sequence_length (int): average sequence length excluding paddings
        random_sequence_length (bool): whether use uniform random number for sequence length
        mask_type (int): mask type - 1: mask index (sequence length excluding paddings). Shape is (batch_size).
                                     2: 2D attention mask. Shape is (batch_size, sequence_length).
                                     3: key len, cumulated lengths of query and key. Shape is (3 * batch_size + 2).

    Returns:
        np.ndarray: the input tensor created
    """

    assert input_mask.type.tensor_type.elem_type in [
        TensorProto.FLOAT,
        TensorProto.INT32,
        TensorProto.INT64,
    ]

    if mask_type == 1:  # sequence length excluding paddings
        data = np.ones((batch_size), dtype=np.int32)
        if random_sequence_length:
            for i in range(batch_size):
                data[i] = get_random_length(sequence_length, average_sequence_length)
        else:
            for i in range(batch_size):
                data[i] = average_sequence_length
    elif mask_type == 2:  # 2D attention mask
        data = np.zeros((batch_size, sequence_length), dtype=np.int32)
        if random_sequence_length:
            for i in range(batch_size):
                actual_seq_len = get_random_length(sequence_length, average_sequence_length)
                for j in range(actual_seq_len):
                    data[i, j] = 1
        else:
            temp = np.ones((batch_size, average_sequence_length), dtype=np.int32)
            data[: temp.shape[0], : temp.shape[1]] = temp
    else:
        assert mask_type == 3
        data = np.zeros((batch_size * 3 + 2), dtype=np.int32)
        if random_sequence_length:
            for i in range(batch_size):
                data[i] = get_random_length(sequence_length, average_sequence_length)

            for i in range(batch_size + 1):
                data[batch_size + i] = data[batch_size + i - 1] + data[i - 1] if i > 0 else 0
                data[2 * batch_size + 1 + i] = data[batch_size + i - 1] + data[i - 1] if i > 0 else 0
        else:
            for i in range(batch_size):
                data[i] = average_sequence_length
            for i in range(batch_size + 1):
                data[batch_size + i] = i * average_sequence_length
                data[2 * batch_size + 1 + i] = i * average_sequence_length

    if input_mask.type.tensor_type.elem_type == TensorProto.FLOAT:
        data = np.float32(data)
    elif input_mask.type.tensor_type.elem_type == TensorProto.INT64:
        data = np.int64(data)

    return data


def output_test_data(directory: str, inputs: dict[str, np.ndarray]):
    """Output input tensors of test data to a directory

    Args:
        directory (str): path of a directory
        inputs (Dict[str, np.ndarray]): map from input name to value
    """
    if not os.path.exists(directory):
        try:
            os.mkdir(directory)
        except OSError:
            print(f"Creation of the directory {directory} failed")
        else:
            print(f"Successfully created the directory {directory} ")
    else:
        print(f"Warning: directory {directory} existed. Files will be overwritten.")

    for index, (name, data) in enumerate(inputs.items()):
        tensor = numpy_helper.from_array(data, name)
        with open(os.path.join(directory, f"input_{index}.pb"), "wb") as file:
            file.write(tensor.SerializeToString())


def fake_test_data(
    batch_size: int,
    sequence_length: int,
    test_cases: int,
    dictionary_size: int,
    verbose: bool,
    random_seed: int,
    input_ids: TensorProto,
    segment_ids: TensorProto,
    input_mask: TensorProto,
    average_sequence_length: int,
    random_sequence_length: bool,
    mask_type: int,
):
    """Create given number of input data for testing

    Args:
        batch_size (int): batch size
        sequence_length (int): sequence length
        test_cases (int): number of test cases
        dictionary_size (int): vocabulary size of dictionary for input_ids
        verbose (bool): print more information or not
        random_seed (int): random seed
        input_ids (TensorProto): graph input of input IDs
        segment_ids (TensorProto): graph input of token type IDs
        input_mask (TensorProto): graph input of attention mask
        average_sequence_length (int): average sequence length excluding paddings
        random_sequence_length (bool): whether use uniform random number for sequence length
        mask_type (int): mask type 1 is mask index; 2 is 2D mask; 3 is key len, cumulated lengths of query and key

    Returns:
        List[Dict[str,numpy.ndarray]]: list of test cases, where each test case is a dictionary
                                       with input name as key and a tensor as value
    """
    assert input_ids is not None

    np.random.seed(random_seed)
    random.seed(random_seed)

    all_inputs = []
    for _test_case in range(test_cases):
        input_1 = fake_input_ids_data(input_ids, batch_size, sequence_length, dictionary_size)
        inputs = {input_ids.name: input_1}

        if segment_ids:
            inputs[segment_ids.name] = fake_segment_ids_data(segment_ids, batch_size, sequence_length)

        if input_mask:
            inputs[input_mask.name] = fake_input_mask_data(
                input_mask, batch_size, sequence_length, average_sequence_length, random_sequence_length, mask_type
            )

        if verbose and len(all_inputs) == 0:
            print("Example inputs", inputs)
        all_inputs.append(inputs)
    return all_inputs


def generate_test_data(
    batch_size: int,
    sequence_length: int,
    test_cases: int,
    seed: int,
    verbose: bool,
    input_ids: TensorProto,
    segment_ids: TensorProto,
    input_mask: TensorProto,
    average_sequence_length: int,
    random_sequence_length: bool,
    mask_type: int,
    dictionary_size: int = 10000,
):
    """Create given number of input data for testing

    Args:
        batch_size (int): batch size
        sequence_length (int): sequence length
        test_cases (int): number of test cases
        seed (int): random seed
        verbose (bool): print more information or not
        input_ids (TensorProto): graph input of input IDs
        segment_ids (TensorProto): graph input of token type IDs
        input_mask (TensorProto): graph input of attention mask
        average_sequence_length (int): average sequence length excluding paddings
        random_sequence_length (bool): whether use uniform random number for sequence length
        mask_type (int): mask type 1 is mask index; 2 is 2D mask; 3 is key len, cumulated lengths of query and key

    Returns:
        List[Dict[str,numpy.ndarray]]: list of test cases, where each test case is a dictionary
                                       with input name as key and a tensor as value
    """
    all_inputs = fake_test_data(
        batch_size,
        sequence_length,
        test_cases,
        dictionary_size,
        verbose,
        seed,
        input_ids,
        segment_ids,
        input_mask,
        average_sequence_length,
        random_sequence_length,
        mask_type,
    )
    if len(all_inputs) != test_cases:
        print("Failed to create test data for test.")
    return all_inputs


def get_graph_input_from_embed_node(onnx_model, embed_node, input_index):
    if input_index >= len(embed_node.input):
        return None

    input = embed_node.input[input_index]
    graph_input = onnx_model.find_graph_input(input)
    if graph_input is None:
        parent_node = onnx_model.get_parent(embed_node, input_index)
        if parent_node is not None and parent_node.op_type == "Cast":
            graph_input = onnx_model.find_graph_input(parent_node.input[0])
    return graph_input


def find_bert_inputs(
    onnx_model: OnnxModel,
    input_ids_name: str | None = None,
    segment_ids_name: str | None = None,
    input_mask_name: str | None = None,
) -> tuple[np.ndarray | None, np.ndarray | None, np.ndarray | None]:
    """Find graph inputs for BERT model.
    First, we will deduce inputs from EmbedLayerNormalization node.
    If not found, we will guess the meaning of graph inputs based on naming.

    Args:
        onnx_model (OnnxModel): onnx model object
        input_ids_name (str, optional): Name of graph input for input IDs. Defaults to None.
        segment_ids_name (str, optional): Name of graph input for segment IDs. Defaults to None.
        input_mask_name (str, optional): Name of graph input for attention mask. Defaults to None.

    Raises:
        ValueError: Graph does not have input named of input_ids_name or segment_ids_name or input_mask_name
        ValueError: Expected graph input number does not match with specified input_ids_name, segment_ids_name
                    and input_mask_name

    Returns:
        Tuple[Optional[np.ndarray], Optional[np.ndarray], Optional[np.ndarray]]: input tensors of input_ids,
                                                                                 segment_ids and input_mask
    """

    graph_inputs = onnx_model.get_graph_inputs_excluding_initializers()

    if input_ids_name is not None:
        input_ids = onnx_model.find_graph_input(input_ids_name)
        if input_ids is None:
            raise ValueError(f"Graph does not have input named {input_ids_name}")

        segment_ids = None
        if segment_ids_name:
            segment_ids = onnx_model.find_graph_input(segment_ids_name)
            if segment_ids is None:
                raise ValueError(f"Graph does not have input named {segment_ids_name}")

        input_mask = None
        if input_mask_name:
            input_mask = onnx_model.find_graph_input(input_mask_name)
            if input_mask is None:
                raise ValueError(f"Graph does not have input named {input_mask_name}")

        expected_inputs = 1 + (1 if segment_ids else 0) + (1 if input_mask else 0)
        if len(graph_inputs) != expected_inputs:
            raise ValueError(f"Expect the graph to have {expected_inputs} inputs. Got {len(graph_inputs)}")

        return input_ids, segment_ids, input_mask

    if len(graph_inputs) != 3:
        raise ValueError(f"Expect the graph to have 3 inputs. Got {len(graph_inputs)}")

    embed_nodes = onnx_model.get_nodes_by_op_type("EmbedLayerNormalization")
    if len(embed_nodes) == 1:
        embed_node = embed_nodes[0]
        input_ids = get_graph_input_from_embed_node(onnx_model, embed_node, 0)
        segment_ids = get_graph_input_from_embed_node(onnx_model, embed_node, 1)
        input_mask = get_graph_input_from_embed_node(onnx_model, embed_node, 7)

        if input_mask is None:
            for input in graph_inputs:
                input_name_lower = input.name.lower()
                if "mask" in input_name_lower:
                    input_mask = input
        if input_mask is None:
            raise ValueError("Failed to find attention mask input")

        return input_ids, segment_ids, input_mask

    # Try guess the inputs based on naming.
    input_ids = None
    segment_ids = None
    input_mask = None
    for input in graph_inputs:
        input_name_lower = input.name.lower()
        if "mask" in input_name_lower:  # matches input with name like "attention_mask" or "input_mask"
            input_mask = input
        elif (
            "token" in input_name_lower or "segment" in input_name_lower
        ):  # matches input with name like "segment_ids" or "token_type_ids"
            segment_ids = input
        else:
            input_ids = input

    if input_ids and segment_ids and input_mask:
        return input_ids, segment_ids, input_mask

    raise ValueError("Fail to assign 3 inputs. You might try rename the graph inputs.")


def get_bert_inputs(
    onnx_file: str,
    input_ids_name: str | None = None,
    segment_ids_name: str | None = None,
    input_mask_name: str | None = None,
) -> tuple[np.ndarray | None, np.ndarray | None, np.ndarray | None]:
    """Find graph inputs for BERT model.
    First, we will deduce inputs from EmbedLayerNormalization node.
    If not found, we will guess the meaning of graph inputs based on naming.

    Args:
        onnx_file (str): onnx model path
        input_ids_name (str, optional): Name of graph input for input IDs. Defaults to None.
        segment_ids_name (str, optional): Name of graph input for segment IDs. Defaults to None.
        input_mask_name (str, optional): Name of graph input for attention mask. Defaults to None.

    Returns:
        Tuple[Optional[np.ndarray], Optional[np.ndarray], Optional[np.ndarray]]: input tensors of input_ids,
                                                                                 segment_ids and input_mask
    """
    model = ModelProto()
    with open(onnx_file, "rb") as file:
        model.ParseFromString(file.read())

    onnx_model = OnnxModel(model)
    return find_bert_inputs(onnx_model, input_ids_name, segment_ids_name, input_mask_name)


def parse_arguments():
    parser = argparse.ArgumentParser()

    parser.add_argument("--model", required=True, type=str, help="bert onnx model path.")

    parser.add_argument(
        "--output_dir",
        required=False,
        type=str,
        default=None,
        help="output test data path. Default is current directory.",
    )

    parser.add_argument("--batch_size", required=False, type=int, default=1, help="batch size of input")

    parser.add_argument(
        "--sequence_length",
        required=False,
        type=int,
        default=128,
        help="maximum sequence length of input",
    )

    parser.add_argument(
        "--input_ids_name",
        required=False,
        type=str,
        default=None,
        help="input name for input ids",
    )
    parser.add_argument(
        "--segment_ids_name",
        required=False,
        type=str,
        default=None,
        help="input name for segment ids",
    )
    parser.add_argument(
        "--input_mask_name",
        required=False,
        type=str,
        default=None,
        help="input name for attention mask",
    )

    parser.add_argument(
        "--samples",
        required=False,
        type=int,
        default=1,
        help="number of test cases to be generated",
    )

    parser.add_argument("--seed", required=False, type=int, default=3, help="random seed")

    parser.add_argument(
        "--verbose",
        required=False,
        action="store_true",
        help="print verbose information",
    )
    parser.set_defaults(verbose=False)

    parser.add_argument(
        "--only_input_tensors",
        required=False,
        action="store_true",
        help="only save input tensors and no output tensors",
    )
    parser.set_defaults(only_input_tensors=False)

    parser.add_argument(
        "-a",
        "--average_sequence_length",
        default=-1,
        type=int,
        help="average sequence length excluding padding",
    )

    parser.add_argument(
        "-r",
        "--random_sequence_length",
        required=False,
        action="store_true",
        help="use uniform random instead of fixed sequence length",
    )
    parser.set_defaults(random_sequence_length=False)

    parser.add_argument(
        "--mask_type",
        required=False,
        type=int,
        default=2,
        help="mask type: (1: mask index, 2: raw 2D mask, 3: key lengths, cumulated lengths of query and key)",
    )

    args = parser.parse_args()
    return args


def create_and_save_test_data(
    model: str,
    output_dir: str,
    batch_size: int,
    sequence_length: int,
    test_cases: int,
    seed: int,
    verbose: bool,
    input_ids_name: str | None,
    segment_ids_name: str | None,
    input_mask_name: str | None,
    only_input_tensors: bool,
    average_sequence_length: int,
    random_sequence_length: bool,
    mask_type: int,
):
    """Create test data for a model, and save test data to a directory.

    Args:
        model (str): path of ONNX bert model
        output_dir (str): output directory
        batch_size (int): batch size
        sequence_length (int): sequence length
        test_cases (int): number of test cases
        seed (int): random seed
        verbose (bool): whether print more information
        input_ids_name (str): graph input name of input_ids
        segment_ids_name (str): graph input name of segment_ids
        input_mask_name (str): graph input name of input_mask
        only_input_tensors (bool): only save input tensors,
        average_sequence_length (int): average sequence length excluding paddings
        random_sequence_length (bool): whether use uniform random number for sequence length
        mask_type(int): mask type
    """
    input_ids, segment_ids, input_mask = get_bert_inputs(model, input_ids_name, segment_ids_name, input_mask_name)

    all_inputs = generate_test_data(
        batch_size,
        sequence_length,
        test_cases,
        seed,
        verbose,
        input_ids,
        segment_ids,
        input_mask,
        average_sequence_length,
        random_sequence_length,
        mask_type,
    )

    for i, inputs in enumerate(all_inputs):
        directory = os.path.join(output_dir, "test_data_set_" + str(i))
        output_test_data(directory, inputs)

    if only_input_tensors:
        return

    import onnxruntime

    providers = (
        ["CUDAExecutionProvider", "CPUExecutionProvider"]
        if "CUDAExecutionProvider" in onnxruntime.get_available_providers()
        else ["CPUExecutionProvider"]
    )
    session = onnxruntime.InferenceSession(model, providers=providers)
    output_names = [output.name for output in session.get_outputs()]

    for i, inputs in enumerate(all_inputs):
        directory = os.path.join(output_dir, "test_data_set_" + str(i))
        result = session.run(output_names, inputs)
        for i, output_name in enumerate(output_names):  # noqa: PLW2901
            tensor_result = numpy_helper.from_array(np.asarray(result[i]), output_name)
            with open(os.path.join(directory, f"output_{i}.pb"), "wb") as file:
                file.write(tensor_result.SerializeToString())


def main():
    args = parse_arguments()

    if args.average_sequence_length <= 0:
        args.average_sequence_length = args.sequence_length

    output_dir = args.output_dir
    if output_dir is None:
        # Default output directory is a sub-directory under the directory of model.
        p = Path(args.model)
        output_dir = os.path.join(p.parent, f"batch_{args.batch_size}_seq_{args.sequence_length}")

    if output_dir is not None:
        # create the output directory if not existed
        path = Path(output_dir)
        path.mkdir(parents=True, exist_ok=True)
    else:
        print("Directory existed. test data files will be overwritten.")

    create_and_save_test_data(
        args.model,
        output_dir,
        args.batch_size,
        args.sequence_length,
        args.samples,
        args.seed,
        args.verbose,
        args.input_ids_name,
        args.segment_ids_name,
        args.input_mask_name,
        args.only_input_tensors,
        args.average_sequence_length,
        args.random_sequence_length,
        args.mask_type,
    )

    print("Test data is saved to directory:", output_dir)


if __name__ == "__main__":
    main()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\geometry\entity.py ===
"""The definition of the base geometrical entity with attributes common to
all derived geometrical entities.

Contains
========

GeometryEntity
GeometricSet

Notes
=====

A GeometryEntity is any object that has special geometric properties.
A GeometrySet is a superclass of any GeometryEntity that can also
be viewed as a sympy.sets.Set.  In particular, points are the only
GeometryEntity not considered a Set.

Rn is a GeometrySet representing n-dimensional Euclidean space. R2 and
R3 are currently the only ambient spaces implemented.

"""
from __future__ import annotations

from sympy.core.basic import Basic
from sympy.core.containers import Tuple
from sympy.core.evalf import EvalfMixin, N
from sympy.core.numbers import oo
from sympy.core.symbol import Dummy
from sympy.core.sympify import sympify
from sympy.functions.elementary.trigonometric import cos, sin, atan
from sympy.matrices import eye
from sympy.multipledispatch import dispatch
from sympy.printing import sstr
from sympy.sets import Set, Union, FiniteSet
from sympy.sets.handlers.intersection import intersection_sets
from sympy.sets.handlers.union import union_sets
from sympy.solvers.solvers import solve
from sympy.utilities.misc import func_name
from sympy.utilities.iterables import is_sequence


# How entities are ordered; used by __cmp__ in GeometryEntity
ordering_of_classes = [
    "Point2D",
    "Point3D",
    "Point",
    "Segment2D",
    "Ray2D",
    "Line2D",
    "Segment3D",
    "Line3D",
    "Ray3D",
    "Segment",
    "Ray",
    "Line",
    "Plane",
    "Triangle",
    "RegularPolygon",
    "Polygon",
    "Circle",
    "Ellipse",
    "Curve",
    "Parabola"
]


x, y = [Dummy('entity_dummy') for i in range(2)]
T = Dummy('entity_dummy', real=True)


class GeometryEntity(Basic, EvalfMixin):
    """The base class for all geometrical entities.

    This class does not represent any particular geometric entity, it only
    provides the implementation of some methods common to all subclasses.

    """

    __slots__: tuple[str, ...] = ()

    def __cmp__(self, other):
        """Comparison of two GeometryEntities."""
        n1 = self.__class__.__name__
        n2 = other.__class__.__name__
        c = (n1 > n2) - (n1 < n2)
        if not c:
            return 0

        i1 = -1
        for cls in self.__class__.__mro__:
            try:
                i1 = ordering_of_classes.index(cls.__name__)
                break
            except ValueError:
                i1 = -1
        if i1 == -1:
            return c

        i2 = -1
        for cls in other.__class__.__mro__:
            try:
                i2 = ordering_of_classes.index(cls.__name__)
                break
            except ValueError:
                i2 = -1
        if i2 == -1:
            return c

        return (i1 > i2) - (i1 < i2)

    def __contains__(self, other):
        """Subclasses should implement this method for anything more complex than equality."""
        if type(self) is type(other):
            return self == other
        raise NotImplementedError()

    def __getnewargs__(self):
        """Returns a tuple that will be passed to __new__ on unpickling."""
        return tuple(self.args)

    def __ne__(self, o):
        """Test inequality of two geometrical entities."""
        return not self == o

    def __new__(cls, *args, **kwargs):
        # Points are sequences, but they should not
        # be converted to Tuples, so use this detection function instead.
        def is_seq_and_not_point(a):
            # we cannot use isinstance(a, Point) since we cannot import Point
            if hasattr(a, 'is_Point') and a.is_Point:
                return False
            return is_sequence(a)

        args = [Tuple(*a) if is_seq_and_not_point(a) else sympify(a) for a in args]
        return Basic.__new__(cls, *args)

    def __radd__(self, a):
        """Implementation of reverse add method."""
        return a.__add__(self)

    def __rtruediv__(self, a):
        """Implementation of reverse division method."""
        return a.__truediv__(self)

    def __repr__(self):
        """String representation of a GeometryEntity that can be evaluated
        by sympy."""
        return type(self).__name__ + repr(self.args)

    def __rmul__(self, a):
        """Implementation of reverse multiplication method."""
        return a.__mul__(self)

    def __rsub__(self, a):
        """Implementation of reverse subtraction method."""
        return a.__sub__(self)

    def __str__(self):
        """String representation of a GeometryEntity."""
        return type(self).__name__ + sstr(self.args)

    def _eval_subs(self, old, new):
        from sympy.geometry.point import Point, Point3D
        if is_sequence(old) or is_sequence(new):
            if isinstance(self, Point3D):
                old = Point3D(old)
                new = Point3D(new)
            else:
                old = Point(old)
                new = Point(new)
            return  self._subs(old, new)

    def _repr_svg_(self):
        """SVG representation of a GeometryEntity suitable for IPython"""

        try:
            bounds = self.bounds
        except (NotImplementedError, TypeError):
            # if we have no SVG representation, return None so IPython
            # will fall back to the next representation
            return None

        if not all(x.is_number and x.is_finite for x in bounds):
            return None

        svg_top = '''<svg xmlns="http://www.w3.org/2000/svg"
            xmlns:xlink="http://www.w3.org/1999/xlink"
            width="{1}" height="{2}" viewBox="{0}"
            preserveAspectRatio="xMinYMin meet">
            <defs>
                <marker id="markerCircle" markerWidth="8" markerHeight="8"
                    refx="5" refy="5" markerUnits="strokeWidth">
                    <circle cx="5" cy="5" r="1.5" style="stroke: none; fill:#000000;"/>
                </marker>
                <marker id="markerArrow" markerWidth="13" markerHeight="13" refx="2" refy="4"
                       orient="auto" markerUnits="strokeWidth">
                    <path d="M2,2 L2,6 L6,4" style="fill: #000000;" />
                </marker>
                <marker id="markerReverseArrow" markerWidth="13" markerHeight="13" refx="6" refy="4"
                       orient="auto" markerUnits="strokeWidth">
                    <path d="M6,2 L6,6 L2,4" style="fill: #000000;" />
                </marker>
            </defs>'''

        # Establish SVG canvas that will fit all the data + small space
        xmin, ymin, xmax, ymax = map(N, bounds)
        if xmin == xmax and ymin == ymax:
            # This is a point; buffer using an arbitrary size
            xmin, ymin, xmax, ymax = xmin - .5, ymin -.5, xmax + .5, ymax + .5
        else:
            # Expand bounds by a fraction of the data ranges
            expand = 0.1  # or 10%; this keeps arrowheads in view (R plots use 4%)
            widest_part = max([xmax - xmin, ymax - ymin])
            expand_amount = widest_part * expand
            xmin -= expand_amount
            ymin -= expand_amount
            xmax += expand_amount
            ymax += expand_amount
        dx = xmax - xmin
        dy = ymax - ymin
        width = min([max([100., dx]), 300])
        height = min([max([100., dy]), 300])

        scale_factor = 1. if max(width, height) == 0 else max(dx, dy) / max(width, height)
        try:
            svg = self._svg(scale_factor)
        except (NotImplementedError, TypeError):
            # if we have no SVG representation, return None so IPython
            # will fall back to the next representation
            return None

        view_box = "{} {} {} {}".format(xmin, ymin, dx, dy)
        transform = "matrix(1,0,0,-1,0,{})".format(ymax + ymin)
        svg_top = svg_top.format(view_box, width, height)

        return svg_top + (
            '<g transform="{}">{}</g></svg>'
            ).format(transform, svg)

    def _svg(self, scale_factor=1., fill_color="#66cc99"):
        """Returns SVG path element for the GeometryEntity.

        Parameters
        ==========

        scale_factor : float
            Multiplication factor for the SVG stroke-width.  Default is 1.
        fill_color : str, optional
            Hex string for fill color. Default is "#66cc99".
        """
        raise NotImplementedError()

    def _sympy_(self):
        return self

    @property
    def ambient_dimension(self):
        """What is the dimension of the space that the object is contained in?"""
        raise NotImplementedError()

    @property
    def bounds(self):
        """Return a tuple (xmin, ymin, xmax, ymax) representing the bounding
        rectangle for the geometric figure.

        """

        raise NotImplementedError()

    def encloses(self, o):
        """
        Return True if o is inside (not on or outside) the boundaries of self.

        The object will be decomposed into Points and individual Entities need
        only define an encloses_point method for their class.

        See Also
        ========

        sympy.geometry.ellipse.Ellipse.encloses_point
        sympy.geometry.polygon.Polygon.encloses_point

        Examples
        ========

        >>> from sympy import RegularPolygon, Point, Polygon
        >>> t  = Polygon(*RegularPolygon(Point(0, 0), 1, 3).vertices)
        >>> t2 = Polygon(*RegularPolygon(Point(0, 0), 2, 3).vertices)
        >>> t2.encloses(t)
        True
        >>> t.encloses(t2)
        False

        """

        from sympy.geometry.point import Point
        from sympy.geometry.line import Segment, Ray, Line
        from sympy.geometry.ellipse import Ellipse
        from sympy.geometry.polygon import Polygon, RegularPolygon

        if isinstance(o, Point):
            return self.encloses_point(o)
        elif isinstance(o, Segment):
            return all(self.encloses_point(x) for x in o.points)
        elif isinstance(o, (Ray, Line)):
            return False
        elif isinstance(o, Ellipse):
            return self.encloses_point(o.center) and \
                self.encloses_point(
                Point(o.center.x + o.hradius, o.center.y)) and \
                not self.intersection(o)
        elif isinstance(o, Polygon):
            if isinstance(o, RegularPolygon):
                if not self.encloses_point(o.center):
                    return False
            return all(self.encloses_point(v) for v in o.vertices)
        raise NotImplementedError()

    def equals(self, o):
        return self == o

    def intersection(self, o):
        """
        Returns a list of all of the intersections of self with o.

        Notes
        =====

        An entity is not required to implement this method.

        If two different types of entities can intersect, the item with
        higher index in ordering_of_classes should implement
        intersections with anything having a lower index.

        See Also
        ========

        sympy.geometry.util.intersection

        """
        raise NotImplementedError()

    def is_similar(self, other):
        """Is this geometrical entity similar to another geometrical entity?

        Two entities are similar if a uniform scaling (enlarging or
        shrinking) of one of the entities will allow one to obtain the other.

        Notes
        =====

        This method is not intended to be used directly but rather
        through the `are_similar` function found in util.py.
        An entity is not required to implement this method.
        If two different types of entities can be similar, it is only
        required that one of them be able to determine this.

        See Also
        ========

        scale

        """
        raise NotImplementedError()

    def reflect(self, line):
        """
        Reflects an object across a line.

        Parameters
        ==========

        line: Line

        Examples
        ========

        >>> from sympy import pi, sqrt, Line, RegularPolygon
        >>> l = Line((0, pi), slope=sqrt(2))
        >>> pent = RegularPolygon((1, 2), 1, 5)
        >>> rpent = pent.reflect(l)
        >>> rpent
        RegularPolygon(Point2D(-2*sqrt(2)*pi/3 - 1/3 + 4*sqrt(2)/3, 2/3 + 2*sqrt(2)/3 + 2*pi/3), -1, 5, -atan(2*sqrt(2)) + 3*pi/5)

        >>> from sympy import pi, Line, Circle, Point
        >>> l = Line((0, pi), slope=1)
        >>> circ = Circle(Point(0, 0), 5)
        >>> rcirc = circ.reflect(l)
        >>> rcirc
        Circle(Point2D(-pi, pi), -5)

        """
        from sympy.geometry.point import Point

        g = self
        l = line
        o = Point(0, 0)
        if l.slope.is_zero:
            v = l.args[0].y
            if not v:  # x-axis
                return g.scale(y=-1)
            reps = [(p, p.translate(y=2*(v - p.y))) for p in g.atoms(Point)]
        elif l.slope is oo:
            v = l.args[0].x
            if not v:  # y-axis
                return g.scale(x=-1)
            reps = [(p, p.translate(x=2*(v - p.x))) for p in g.atoms(Point)]
        else:
            if not hasattr(g, 'reflect') and not all(
                    isinstance(arg, Point) for arg in g.args):
                raise NotImplementedError(
                    'reflect undefined or non-Point args in %s' % g)
            a = atan(l.slope)
            c = l.coefficients
            d = -c[-1]/c[1]  # y-intercept
            # apply the transform to a single point
            xf = Point(x, y)
            xf = xf.translate(y=-d).rotate(-a, o).scale(y=-1
                ).rotate(a, o).translate(y=d)
            # replace every point using that transform
            reps = [(p, xf.xreplace({x: p.x, y: p.y})) for p in g.atoms(Point)]
        return g.xreplace(dict(reps))

    def rotate(self, angle, pt=None):
        """Rotate ``angle`` radians counterclockwise about Point ``pt``.

        The default pt is the origin, Point(0, 0)

        See Also
        ========

        scale, translate

        Examples
        ========

        >>> from sympy import Point, RegularPolygon, Polygon, pi
        >>> t = Polygon(*RegularPolygon(Point(0, 0), 1, 3).vertices)
        >>> t # vertex on x axis
        Triangle(Point2D(1, 0), Point2D(-1/2, sqrt(3)/2), Point2D(-1/2, -sqrt(3)/2))
        >>> t.rotate(pi/2) # vertex on y axis now
        Triangle(Point2D(0, 1), Point2D(-sqrt(3)/2, -1/2), Point2D(sqrt(3)/2, -1/2))

        """
        newargs = []
        for a in self.args:
            if isinstance(a, GeometryEntity):
                newargs.append(a.rotate(angle, pt))
            else:
                newargs.append(a)
        return type(self)(*newargs)

    def scale(self, x=1, y=1, pt=None):
        """Scale the object by multiplying the x,y-coordinates by x and y.

        If pt is given, the scaling is done relative to that point; the
        object is shifted by -pt, scaled, and shifted by pt.

        See Also
        ========

        rotate, translate

        Examples
        ========

        >>> from sympy import RegularPolygon, Point, Polygon
        >>> t = Polygon(*RegularPolygon(Point(0, 0), 1, 3).vertices)
        >>> t
        Triangle(Point2D(1, 0), Point2D(-1/2, sqrt(3)/2), Point2D(-1/2, -sqrt(3)/2))
        >>> t.scale(2)
        Triangle(Point2D(2, 0), Point2D(-1, sqrt(3)/2), Point2D(-1, -sqrt(3)/2))
        >>> t.scale(2, 2)
        Triangle(Point2D(2, 0), Point2D(-1, sqrt(3)), Point2D(-1, -sqrt(3)))

        """
        from sympy.geometry.point import Point
        if pt:
            pt = Point(pt, dim=2)
            return self.translate(*(-pt).args).scale(x, y).translate(*pt.args)
        return type(self)(*[a.scale(x, y) for a in self.args])  # if this fails, override this class

    def translate(self, x=0, y=0):
        """Shift the object by adding to the x,y-coordinates the values x and y.

        See Also
        ========

        rotate, scale

        Examples
        ========

        >>> from sympy import RegularPolygon, Point, Polygon
        >>> t = Polygon(*RegularPolygon(Point(0, 0), 1, 3).vertices)
        >>> t
        Triangle(Point2D(1, 0), Point2D(-1/2, sqrt(3)/2), Point2D(-1/2, -sqrt(3)/2))
        >>> t.translate(2)
        Triangle(Point2D(3, 0), Point2D(3/2, sqrt(3)/2), Point2D(3/2, -sqrt(3)/2))
        >>> t.translate(2, 2)
        Triangle(Point2D(3, 2), Point2D(3/2, sqrt(3)/2 + 2), Point2D(3/2, 2 - sqrt(3)/2))

        """
        newargs = []
        for a in self.args:
            if isinstance(a, GeometryEntity):
                newargs.append(a.translate(x, y))
            else:
                newargs.append(a)
        return self.func(*newargs)

    def parameter_value(self, other, t):
        """Return the parameter corresponding to the given point.
        Evaluating an arbitrary point of the entity at this parameter
        value will return the given point.

        Examples
        ========

        >>> from sympy import Line, Point
        >>> from sympy.abc import t
        >>> a = Point(0, 0)
        >>> b = Point(2, 2)
        >>> Line(a, b).parameter_value((1, 1), t)
        {t: 1/2}
        >>> Line(a, b).arbitrary_point(t).subs(_)
        Point2D(1, 1)
        """
        from sympy.geometry.point import Point
        if not isinstance(other, GeometryEntity):
            other = Point(other, dim=self.ambient_dimension)
        if not isinstance(other, Point):
            raise ValueError("other must be a point")
        sol = solve(self.arbitrary_point(T) - other, T, dict=True)
        if not sol:
            raise ValueError("Given point is not on %s" % func_name(self))
        return {t: sol[0][T]}


class GeometrySet(GeometryEntity, Set):
    """Parent class of all GeometryEntity that are also Sets
    (compatible with sympy.sets)
    """
    __slots__ = ()

    def _contains(self, other):
        """sympy.sets uses the _contains method, so include it for compatibility."""

        if isinstance(other, Set) and other.is_FiniteSet:
            return all(self.__contains__(i) for i in other)

        return self.__contains__(other)

@dispatch(GeometrySet, Set)  # type:ignore # noqa:F811
def union_sets(self, o): # noqa:F811
    """ Returns the union of self and o
    for use with sympy.sets.Set, if possible. """


    # if its a FiniteSet, merge any points
    # we contain and return a union with the rest
    if o.is_FiniteSet:
        other_points = [p for p in o if not self._contains(p)]
        if len(other_points) == len(o):
            return None
        return Union(self, FiniteSet(*other_points))
    if self._contains(o):
        return self
    return None


@dispatch(GeometrySet, Set)  # type: ignore # noqa:F811
def intersection_sets(self, o): # noqa:F811
    """ Returns a sympy.sets.Set of intersection objects,
    if possible. """

    from sympy.geometry.point import Point

    try:
        # if o is a FiniteSet, find the intersection directly
        # to avoid infinite recursion
        if o.is_FiniteSet:
            inter = FiniteSet(*(p for p in o if self.contains(p)))
        else:
            inter = self.intersection(o)
    except NotImplementedError:
        # sympy.sets.Set.reduce expects None if an object
        # doesn't know how to simplify
        return None

    # put the points in a FiniteSet
    points = FiniteSet(*[p for p in inter if isinstance(p, Point)])
    non_points = [p for p in inter if not isinstance(p, Point)]

    return Union(*(non_points + [points]))

def translate(x, y):
    """Return the matrix to translate a 2-D point by x and y."""
    rv = eye(3)
    rv[2, 0] = x
    rv[2, 1] = y
    return rv


def scale(x, y, pt=None):
    """Return the matrix to multiply a 2-D point's coordinates by x and y.

    If pt is given, the scaling is done relative to that point."""
    rv = eye(3)
    rv[0, 0] = x
    rv[1, 1] = y
    if pt:
        from sympy.geometry.point import Point
        pt = Point(pt, dim=2)
        tr1 = translate(*(-pt).args)
        tr2 = translate(*pt.args)
        return tr1*rv*tr2
    return rv


def rotate(th):
    """Return the matrix to rotate a 2-D point about the origin by ``angle``.

    The angle is measured in radians. To Point a point about a point other
    then the origin, translate the Point, do the rotation, and
    translate it back:

    >>> from sympy.geometry.entity import rotate, translate
    >>> from sympy import Point, pi
    >>> rot_about_11 = translate(-1, -1)*rotate(pi/2)*translate(1, 1)
    >>> Point(1, 1).transform(rot_about_11)
    Point2D(1, 1)
    >>> Point(0, 0).transform(rot_about_11)
    Point2D(2, 0)
    """
    s = sin(th)
    rv = eye(3)*cos(th)
    rv[0, 1] = s
    rv[1, 0] = -s
    rv[2, 2] = 1
    return rv

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\mechanics\wrapping_geometry.py ===
"""Geometry objects for use by wrapping pathways."""

from abc import ABC, abstractmethod

from sympy import Integer, acos, pi, sqrt, sympify, tan
from sympy.core.relational import Eq
from sympy.functions.elementary.trigonometric import atan2
from sympy.polys.polytools import cancel
from sympy.physics.vector import Vector, dot
from sympy.simplify.simplify import trigsimp


__all__ = [
    'WrappingGeometryBase',
    'WrappingCylinder',
    'WrappingSphere',
]


class WrappingGeometryBase(ABC):
    """Abstract base class for all geometry classes to inherit from.

    Notes
    =====

    Instances of this class cannot be directly instantiated by users. However,
    it can be used to created custom geometry types through subclassing.

    """

    @property
    @abstractmethod
    def point(cls):
        """The point with which the geometry is associated."""
        pass

    @abstractmethod
    def point_on_surface(self, point):
        """Returns ``True`` if a point is on the geometry's surface.

        Parameters
        ==========
        point : Point
            The point for which it's to be ascertained if it's on the
            geometry's surface or not.

        """
        pass

    @abstractmethod
    def geodesic_length(self, point_1, point_2):
        """Returns the shortest distance between two points on a geometry's
        surface.

        Parameters
        ==========

        point_1 : Point
            The point from which the geodesic length should be calculated.
        point_2 : Point
            The point to which the geodesic length should be calculated.

        """
        pass

    @abstractmethod
    def geodesic_end_vectors(self, point_1, point_2):
        """The vectors parallel to the geodesic at the two end points.

        Parameters
        ==========

        point_1 : Point
            The point from which the geodesic originates.
        point_2 : Point
            The point at which the geodesic terminates.

        """
        pass

    def __repr__(self):
        """Default representation of a geometry object."""
        return f'{self.__class__.__name__}()'


class WrappingSphere(WrappingGeometryBase):
    """A solid spherical object.

    Explanation
    ===========

    A wrapping geometry that allows for circular arcs to be defined between
    pairs of points. These paths are always geodetic (the shortest possible).

    Examples
    ========

    To create a ``WrappingSphere`` instance, a ``Symbol`` denoting its radius
    and ``Point`` at which its center will be located are needed:

    >>> from sympy import symbols
    >>> from sympy.physics.mechanics import Point, WrappingSphere
    >>> r = symbols('r')
    >>> pO = Point('pO')

    A sphere with radius ``r`` centered on ``pO`` can be instantiated with:

    >>> WrappingSphere(r, pO)
    WrappingSphere(radius=r, point=pO)

    Parameters
    ==========

    radius : Symbol
        Radius of the sphere. This symbol must represent a value that is
        positive and constant, i.e. it cannot be a dynamic symbol, nor can it
        be an expression.
    point : Point
        A point at which the sphere is centered.

    See Also
    ========

    WrappingCylinder: Cylindrical geometry where the wrapping direction can be
        defined.

    """

    def __init__(self, radius, point):
        """Initializer for ``WrappingSphere``.

        Parameters
        ==========

        radius : Symbol
            The radius of the sphere.
        point : Point
            A point on which the sphere is centered.

        """
        self.radius = radius
        self.point = point

    @property
    def radius(self):
        """Radius of the sphere."""
        return self._radius

    @radius.setter
    def radius(self, radius):
        self._radius = radius

    @property
    def point(self):
        """A point on which the sphere is centered."""
        return self._point

    @point.setter
    def point(self, point):
        self._point = point

    def point_on_surface(self, point):
        """Returns ``True`` if a point is on the sphere's surface.

        Parameters
        ==========

        point : Point
            The point for which it's to be ascertained if it's on the sphere's
            surface or not. This point's position relative to the sphere's
            center must be a simple expression involving the radius of the
            sphere, otherwise this check will likely not work.

        """
        point_vector = point.pos_from(self.point)
        if isinstance(point_vector, Vector):
            point_radius_squared = dot(point_vector, point_vector)
        else:
            point_radius_squared = point_vector**2
        return Eq(point_radius_squared, self.radius**2) == True

    def geodesic_length(self, point_1, point_2):
        r"""Returns the shortest distance between two points on the sphere's
        surface.

        Explanation
        ===========

        The geodesic length, i.e. the shortest arc along the surface of a
        sphere, connecting two points can be calculated using the formula:

        .. math::

           l = \arccos\left(\mathbf{v}_1 \cdot \mathbf{v}_2\right)

        where $\mathbf{v}_1$ and $\mathbf{v}_2$ are the unit vectors from the
        sphere's center to the first and second points on the sphere's surface
        respectively. Note that the actual path that the geodesic will take is
        undefined when the two points are directly opposite one another.

        Examples
        ========

        A geodesic length can only be calculated between two points on the
        sphere's surface. Firstly, a ``WrappingSphere`` instance must be
        created along with two points that will lie on its surface:

        >>> from sympy import symbols
        >>> from sympy.physics.mechanics import (Point, ReferenceFrame,
        ...     WrappingSphere)
        >>> N = ReferenceFrame('N')
        >>> r = symbols('r')
        >>> pO = Point('pO')
        >>> pO.set_vel(N, 0)
        >>> sphere = WrappingSphere(r, pO)
        >>> p1 = Point('p1')
        >>> p2 = Point('p2')

        Let's assume that ``p1`` lies at a distance of ``r`` in the ``N.x``
        direction from ``pO`` and that ``p2`` is located on the sphere's
        surface in the ``N.y + N.z`` direction from ``pO``. These positions can
        be set with:

        >>> p1.set_pos(pO, r*N.x)
        >>> p1.pos_from(pO)
        r*N.x
        >>> p2.set_pos(pO, r*(N.y + N.z).normalize())
        >>> p2.pos_from(pO)
        sqrt(2)*r/2*N.y + sqrt(2)*r/2*N.z

        The geodesic length, which is in this case is a quarter of the sphere's
        circumference, can be calculated using the ``geodesic_length`` method:

        >>> sphere.geodesic_length(p1, p2)
        pi*r/2

        If the ``geodesic_length`` method is passed an argument, the ``Point``
        that doesn't lie on the sphere's surface then a ``ValueError`` is
        raised because it's not possible to calculate a value in this case.

        Parameters
        ==========

        point_1 : Point
            Point from which the geodesic length should be calculated.
        point_2 : Point
            Point to which the geodesic length should be calculated.

        """
        for point in (point_1, point_2):
            if not self.point_on_surface(point):
                msg = (
                    f'Geodesic length cannot be calculated as point {point} '
                    f'with radius {point.pos_from(self.point).magnitude()} '
                    f'from the sphere\'s center {self.point} does not lie on '
                    f'the surface of {self} with radius {self.radius}.'
                )
                raise ValueError(msg)
        point_1_vector = point_1.pos_from(self.point).normalize()
        point_2_vector = point_2.pos_from(self.point).normalize()
        central_angle = acos(point_2_vector.dot(point_1_vector))
        geodesic_length = self.radius*central_angle
        return geodesic_length

    def geodesic_end_vectors(self, point_1, point_2):
        """The vectors parallel to the geodesic at the two end points.

        Parameters
        ==========

        point_1 : Point
            The point from which the geodesic originates.
        point_2 : Point
            The point at which the geodesic terminates.

        """
        pA, pB = point_1, point_2
        pO = self.point
        pA_vec = pA.pos_from(pO)
        pB_vec = pB.pos_from(pO)

        if pA_vec.cross(pB_vec) == 0:
            msg = (
                f'Can\'t compute geodesic end vectors for the pair of points '
                f'{pA} and {pB} on a sphere {self} as they are diametrically '
                f'opposed, thus the geodesic is not defined.'
            )
            raise ValueError(msg)

        return (
            pA_vec.cross(pB.pos_from(pA)).cross(pA_vec).normalize(),
            pB_vec.cross(pA.pos_from(pB)).cross(pB_vec).normalize(),
        )

    def __repr__(self):
        """Representation of a ``WrappingSphere``."""
        return (
            f'{self.__class__.__name__}(radius={self.radius}, '
            f'point={self.point})'
        )


class WrappingCylinder(WrappingGeometryBase):
    """A solid (infinite) cylindrical object.

    Explanation
    ===========

    A wrapping geometry that allows for circular arcs to be defined between
    pairs of points. These paths are always geodetic (the shortest possible) in
    the sense that they will be a straight line on the unwrapped cylinder's
    surface. However, it is also possible for a direction to be specified, i.e.
    paths can be influenced such that they either wrap along the shortest side
    or the longest side of the cylinder. To define these directions, rotations
    are in the positive direction following the right-hand rule.

    Examples
    ========

    To create a ``WrappingCylinder`` instance, a ``Symbol`` denoting its
    radius, a ``Vector`` defining its axis, and a ``Point`` through which its
    axis passes are needed:

    >>> from sympy import symbols
    >>> from sympy.physics.mechanics import (Point, ReferenceFrame,
    ...     WrappingCylinder)
    >>> N = ReferenceFrame('N')
    >>> r = symbols('r')
    >>> pO = Point('pO')
    >>> ax = N.x

    A cylinder with radius ``r``, and axis parallel to ``N.x`` passing through
    ``pO`` can be instantiated with:

    >>> WrappingCylinder(r, pO, ax)
    WrappingCylinder(radius=r, point=pO, axis=N.x)

    Parameters
    ==========

    radius : Symbol
        The radius of the cylinder.
    point : Point
        A point through which the cylinder's axis passes.
    axis : Vector
        The axis along which the cylinder is aligned.

    See Also
    ========

    WrappingSphere: Spherical geometry where the wrapping direction is always
        geodetic.

    """

    def __init__(self, radius, point, axis):
        """Initializer for ``WrappingCylinder``.

        Parameters
        ==========

        radius : Symbol
            The radius of the cylinder. This symbol must represent a value that
            is positive and constant, i.e. it cannot be a dynamic symbol.
        point : Point
            A point through which the cylinder's axis passes.
        axis : Vector
            The axis along which the cylinder is aligned.

        """
        self.radius = radius
        self.point = point
        self.axis = axis

    @property
    def radius(self):
        """Radius of the cylinder."""
        return self._radius

    @radius.setter
    def radius(self, radius):
        self._radius = radius

    @property
    def point(self):
        """A point through which the cylinder's axis passes."""
        return self._point

    @point.setter
    def point(self, point):
        self._point = point

    @property
    def axis(self):
        """Axis along which the cylinder is aligned."""
        return self._axis

    @axis.setter
    def axis(self, axis):
        self._axis = axis.normalize()

    def point_on_surface(self, point):
        """Returns ``True`` if a point is on the cylinder's surface.

        Parameters
        ==========

        point : Point
            The point for which it's to be ascertained if it's on the
            cylinder's surface or not. This point's position relative to the
            cylinder's axis must be a simple expression involving the radius of
            the sphere, otherwise this check will likely not work.

        """
        relative_position = point.pos_from(self.point)
        parallel = relative_position.dot(self.axis) * self.axis
        point_vector = relative_position - parallel
        if isinstance(point_vector, Vector):
            point_radius_squared = dot(point_vector, point_vector)
        else:
            point_radius_squared = point_vector**2
        return Eq(trigsimp(point_radius_squared), self.radius**2) == True

    def geodesic_length(self, point_1, point_2):
        """The shortest distance between two points on a geometry's surface.

        Explanation
        ===========

        The geodesic length, i.e. the shortest arc along the surface of a
        cylinder, connecting two points. It can be calculated using Pythagoras'
        theorem. The first short side is the distance between the two points on
        the cylinder's surface parallel to the cylinder's axis. The second
        short side is the arc of a circle between the two points of the
        cylinder's surface perpendicular to the cylinder's axis. The resulting
        hypotenuse is the geodesic length.

        Examples
        ========

        A geodesic length can only be calculated between two points on the
        cylinder's surface. Firstly, a ``WrappingCylinder`` instance must be
        created along with two points that will lie on its surface:

        >>> from sympy import symbols, cos, sin
        >>> from sympy.physics.mechanics import (Point, ReferenceFrame,
        ...     WrappingCylinder, dynamicsymbols)
        >>> N = ReferenceFrame('N')
        >>> r = symbols('r')
        >>> pO = Point('pO')
        >>> pO.set_vel(N, 0)
        >>> cylinder = WrappingCylinder(r, pO, N.x)
        >>> p1 = Point('p1')
        >>> p2 = Point('p2')

        Let's assume that ``p1`` is located at ``N.x + r*N.y`` relative to
        ``pO`` and that ``p2`` is located at ``r*(cos(q)*N.y + sin(q)*N.z)``
        relative to ``pO``, where ``q(t)`` is a generalized coordinate
        specifying the angle rotated around the ``N.x`` axis according to the
        right-hand rule where ``N.y`` is zero. These positions can be set with:

        >>> q = dynamicsymbols('q')
        >>> p1.set_pos(pO, N.x + r*N.y)
        >>> p1.pos_from(pO)
        N.x + r*N.y
        >>> p2.set_pos(pO, r*(cos(q)*N.y + sin(q)*N.z).normalize())
        >>> p2.pos_from(pO).simplify()
        r*cos(q(t))*N.y + r*sin(q(t))*N.z

        The geodesic length, which is in this case a is the hypotenuse of a
        right triangle where the other two side lengths are ``1`` (parallel to
        the cylinder's axis) and ``r*q(t)`` (parallel to the cylinder's cross
        section), can be calculated using the ``geodesic_length`` method:

        >>> cylinder.geodesic_length(p1, p2).simplify()
        sqrt(r**2*q(t)**2 + 1)

        If the ``geodesic_length`` method is passed an argument ``Point`` that
        doesn't lie on the sphere's surface then a ``ValueError`` is raised
        because it's not possible to calculate a value in this case.

        Parameters
        ==========

        point_1 : Point
            Point from which the geodesic length should be calculated.
        point_2 : Point
            Point to which the geodesic length should be calculated.

        """
        for point in (point_1, point_2):
            if not self.point_on_surface(point):
                msg = (
                    f'Geodesic length cannot be calculated as point {point} '
                    f'with radius {point.pos_from(self.point).magnitude()} '
                    f'from the cylinder\'s center {self.point} does not lie on '
                    f'the surface of {self} with radius {self.radius} and axis '
                    f'{self.axis}.'
                )
                raise ValueError(msg)

        relative_position = point_2.pos_from(point_1)
        parallel_length = relative_position.dot(self.axis)

        point_1_relative_position = point_1.pos_from(self.point)
        point_1_perpendicular_vector = (
            point_1_relative_position
            - point_1_relative_position.dot(self.axis)*self.axis
        ).normalize()

        point_2_relative_position = point_2.pos_from(self.point)
        point_2_perpendicular_vector = (
            point_2_relative_position
            - point_2_relative_position.dot(self.axis)*self.axis
        ).normalize()

        central_angle = _directional_atan(
            cancel(point_1_perpendicular_vector
                .cross(point_2_perpendicular_vector)
                .dot(self.axis)),
            cancel(point_1_perpendicular_vector.dot(point_2_perpendicular_vector)),
        )

        planar_arc_length = self.radius*central_angle
        geodesic_length = sqrt(parallel_length**2 + planar_arc_length**2)
        return geodesic_length

    def geodesic_end_vectors(self, point_1, point_2):
        """The vectors parallel to the geodesic at the two end points.

        Parameters
        ==========

        point_1 : Point
            The point from which the geodesic originates.
        point_2 : Point
            The point at which the geodesic terminates.

        """
        point_1_from_origin_point = point_1.pos_from(self.point)
        point_2_from_origin_point = point_2.pos_from(self.point)

        if point_1_from_origin_point == point_2_from_origin_point:
            msg = (
                f'Cannot compute geodesic end vectors for coincident points '
                f'{point_1} and {point_2} as no geodesic exists.'
            )
            raise ValueError(msg)

        point_1_parallel = point_1_from_origin_point.dot(self.axis) * self.axis
        point_2_parallel = point_2_from_origin_point.dot(self.axis) * self.axis
        point_1_normal = (point_1_from_origin_point - point_1_parallel)
        point_2_normal = (point_2_from_origin_point - point_2_parallel)

        if point_1_normal == point_2_normal:
            point_1_perpendicular = Vector(0)
            point_2_perpendicular = Vector(0)
        else:
            point_1_perpendicular = self.axis.cross(point_1_normal).normalize()
            point_2_perpendicular = -self.axis.cross(point_2_normal).normalize()

        geodesic_length = self.geodesic_length(point_1, point_2)
        relative_position = point_2.pos_from(point_1)
        parallel_length = relative_position.dot(self.axis)
        planar_arc_length = sqrt(geodesic_length**2 - parallel_length**2)

        point_1_vector = (
            planar_arc_length * point_1_perpendicular
            + parallel_length * self.axis
        ).normalize()
        point_2_vector = (
            planar_arc_length * point_2_perpendicular
            - parallel_length * self.axis
        ).normalize()

        return (point_1_vector, point_2_vector)

    def __repr__(self):
        """Representation of a ``WrappingCylinder``."""
        return (
            f'{self.__class__.__name__}(radius={self.radius}, '
            f'point={self.point}, axis={self.axis})'
        )


def _directional_atan(numerator, denominator):
    """Compute atan in a directional sense as required for geodesics.

    Explanation
    ===========

    To be able to control the direction of the geodesic length along the
    surface of a cylinder a dedicated arctangent function is needed that
    properly handles the directionality of different case. This function
    ensures that the central angle is always positive but shifting the case
    where ``atan2`` would return a negative angle to be centered around
    ``2*pi``.

    Notes
    =====

    This function only handles very specific cases, i.e. the ones that are
    expected to be encountered when calculating symbolic geodesics on uniformly
    curved surfaces. As such, ``NotImplemented`` errors can be raised in many
    cases. This function is named with a leader underscore to indicate that it
    only aims to provide very specific functionality within the private scope
    of this module.

    """

    if numerator.is_number and denominator.is_number:
        angle = atan2(numerator, denominator)
        if angle < 0:
            angle += 2 * pi
    elif numerator.is_number:
        msg = (
            f'Cannot compute a directional atan when the numerator {numerator} '
            f'is numeric and the denominator {denominator} is symbolic.'
        )
        raise NotImplementedError(msg)
    elif denominator.is_number:
        msg = (
            f'Cannot compute a directional atan when the numerator {numerator} '
            f'is symbolic and the denominator {denominator} is numeric.'
        )
        raise NotImplementedError(msg)
    else:
        ratio = sympify(trigsimp(numerator / denominator))
        if isinstance(ratio, tan):
            angle = ratio.args[0]
        elif (
            ratio.is_Mul
            and ratio.args[0] == Integer(-1)
            and isinstance(ratio.args[1], tan)
        ):
            angle = 2 * pi - ratio.args[1].args[0]
        else:
            msg = f'Cannot compute a directional atan for the value {ratio}.'
            raise NotImplementedError(msg)

    return angle

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\plotting\experimental_lambdify.py ===
""" rewrite of lambdify - This stuff is not stable at all.

It is for internal use in the new plotting module.
It may (will! see the Q'n'A in the source) be rewritten.

It's completely self contained. Especially it does not use lambdarepr.

It does not aim to replace the current lambdify. Most importantly it will never
ever support anything else than SymPy expressions (no Matrices, dictionaries
and so on).
"""


import re
from sympy.core.numbers import (I, NumberSymbol, oo, zoo)
from sympy.core.symbol import Symbol
from sympy.utilities.iterables import numbered_symbols

#  We parse the expression string into a tree that identifies functions. Then
# we translate the names of the functions and we translate also some strings
# that are not names of functions (all this according to translation
# dictionaries).
#  If the translation goes to another module (like numpy) the
# module is imported and 'func' is translated to 'module.func'.
#  If a function can not be translated, the inner nodes of that part of the
# tree are not translated. So if we have Integral(sqrt(x)), sqrt is not
# translated to np.sqrt and the Integral does not crash.
#  A namespace for all this is generated by crawling the (func, args) tree of
# the expression. The creation of this namespace involves many ugly
# workarounds.
#  The namespace consists of all the names needed for the SymPy expression and
# all the name of modules used for translation. Those modules are imported only
# as a name (import numpy as np) in order to keep the namespace small and
# manageable.

#  Please, if there is a bug, do not try to fix it here! Rewrite this by using
# the method proposed in the last Q'n'A below. That way the new function will
# work just as well, be just as simple, but it wont need any new workarounds.
#  If you insist on fixing it here, look at the workarounds in the function
# sympy_expression_namespace and in lambdify.

# Q: Why are you not using Python abstract syntax tree?
# A: Because it is more complicated and not much more powerful in this case.

# Q: What if I have Symbol('sin') or g=Function('f')?
# A: You will break the algorithm. We should use srepr to defend against this?
#  The problem with Symbol('sin') is that it will be printed as 'sin'. The
# parser will distinguish it from the function 'sin' because functions are
# detected thanks to the opening parenthesis, but the lambda expression won't
# understand the difference if we have also the sin function.
# The solution (complicated) is to use srepr and maybe ast.
#  The problem with the g=Function('f') is that it will be printed as 'f' but in
# the global namespace we have only 'g'. But as the same printer is used in the
# constructor of the namespace there will be no problem.

# Q: What if some of the printers are not printing as expected?
# A: The algorithm wont work. You must use srepr for those cases. But even
# srepr may not print well. All problems with printers should be considered
# bugs.

# Q: What about _imp_ functions?
# A: Those are taken care for by evalf. A special case treatment will work
# faster but it's not worth the code complexity.

# Q: Will ast fix all possible problems?
# A: No. You will always have to use some printer. Even srepr may not work in
# some cases. But if the printer does not work, that should be considered a
# bug.

# Q: Is there same way to fix all possible problems?
# A: Probably by constructing our strings ourself by traversing the (func,
# args) tree and creating the namespace at the same time. That actually sounds
# good.

from sympy.external import import_module
import warnings

#TODO debugging output


class vectorized_lambdify:
    """ Return a sufficiently smart, vectorized and lambdified function.

    Returns only reals.

    Explanation
    ===========

    This function uses experimental_lambdify to created a lambdified
    expression ready to be used with numpy. Many of the functions in SymPy
    are not implemented in numpy so in some cases we resort to Python cmath or
    even to evalf.

    The following translations are tried:
      only numpy complex
      - on errors raised by SymPy trying to work with ndarray:
          only Python cmath and then vectorize complex128

    When using Python cmath there is no need for evalf or float/complex
    because Python cmath calls those.

    This function never tries to mix numpy directly with evalf because numpy
    does not understand SymPy Float. If this is needed one can use the
    float_wrap_evalf/complex_wrap_evalf options of experimental_lambdify or
    better one can be explicit about the dtypes that numpy works with.
    Check numpy bug http://projects.scipy.org/numpy/ticket/1013 to know what
    types of errors to expect.
    """
    def __init__(self, args, expr):
        self.args = args
        self.expr = expr
        self.np = import_module('numpy')

        self.lambda_func_1 = experimental_lambdify(
            args, expr, use_np=True)
        self.vector_func_1 = self.lambda_func_1

        self.lambda_func_2 = experimental_lambdify(
            args, expr, use_python_cmath=True)
        self.vector_func_2 = self.np.vectorize(
            self.lambda_func_2, otypes=[complex])

        self.vector_func = self.vector_func_1
        self.failure = False

    def __call__(self, *args):
        np = self.np

        try:
            temp_args = (np.array(a, dtype=complex) for a in args)
            results = self.vector_func(*temp_args)
            results = np.ma.masked_where(
                np.abs(results.imag) > 1e-7 * np.abs(results),
                results.real, copy=False)
            return results
        except ValueError:
            if self.failure:
                raise

            self.failure = True
            self.vector_func = self.vector_func_2
            warnings.warn(
                'The evaluation of the expression is problematic. '
                'We are trying a failback method that may still work. '
                'Please report this as a bug.')
            return self.__call__(*args)


class lambdify:
    """Returns the lambdified function.

    Explanation
    ===========

    This function uses experimental_lambdify to create a lambdified
    expression. It uses cmath to lambdify the expression. If the function
    is not implemented in Python cmath, Python cmath calls evalf on those
    functions.
    """

    def __init__(self, args, expr):
        self.args = args
        self.expr = expr
        self.lambda_func_1 = experimental_lambdify(
            args, expr, use_python_cmath=True, use_evalf=True)
        self.lambda_func_2 = experimental_lambdify(
            args, expr, use_python_math=True, use_evalf=True)
        self.lambda_func_3 = experimental_lambdify(
            args, expr, use_evalf=True, complex_wrap_evalf=True)
        self.lambda_func = self.lambda_func_1
        self.failure = False

    def __call__(self, args):
        try:
            #The result can be sympy.Float. Hence wrap it with complex type.
            result = complex(self.lambda_func(args))
            if abs(result.imag) > 1e-7 * abs(result):
                return None
            return result.real
        except (ZeroDivisionError, OverflowError):
            return None
        except TypeError as e:
            if self.failure:
                raise e

            if self.lambda_func == self.lambda_func_1:
                self.lambda_func = self.lambda_func_2
                return self.__call__(args)

            self.failure = True
            self.lambda_func = self.lambda_func_3
            warnings.warn(
                'The evaluation of the expression is problematic. '
                'We are trying a failback method that may still work. '
                'Please report this as a bug.', stacklevel=2)
            return self.__call__(args)


def experimental_lambdify(*args, **kwargs):
    l = Lambdifier(*args, **kwargs)
    return l


class Lambdifier:
    def __init__(self, args, expr, print_lambda=False, use_evalf=False,
                 float_wrap_evalf=False, complex_wrap_evalf=False,
                 use_np=False, use_python_math=False, use_python_cmath=False,
                 use_interval=False):

        self.print_lambda = print_lambda
        self.use_evalf = use_evalf
        self.float_wrap_evalf = float_wrap_evalf
        self.complex_wrap_evalf = complex_wrap_evalf
        self.use_np = use_np
        self.use_python_math = use_python_math
        self.use_python_cmath = use_python_cmath
        self.use_interval = use_interval

        # Constructing the argument string
        # - check
        if not all(isinstance(a, Symbol) for a in args):
            raise ValueError('The arguments must be Symbols.')
        # - use numbered symbols
        syms = numbered_symbols(exclude=expr.free_symbols)
        newargs = [next(syms) for _ in args]
        expr = expr.xreplace(dict(zip(args, newargs)))
        argstr = ', '.join([str(a) for a in newargs])
        del syms, newargs, args

        # Constructing the translation dictionaries and making the translation
        self.dict_str = self.get_dict_str()
        self.dict_fun = self.get_dict_fun()
        exprstr = str(expr)
        newexpr = self.tree2str_translate(self.str2tree(exprstr))

        # Constructing the namespaces
        namespace = {}
        namespace.update(self.sympy_atoms_namespace(expr))
        namespace.update(self.sympy_expression_namespace(expr))
        # XXX Workaround
        # Ugly workaround because Pow(a,Half) prints as sqrt(a)
        # and sympy_expression_namespace can not catch it.
        from sympy.functions.elementary.miscellaneous import sqrt
        namespace.update({'sqrt': sqrt})
        namespace.update({'Eq': lambda x, y: x == y})
        namespace.update({'Ne': lambda x, y: x != y})
        # End workaround.
        if use_python_math:
            namespace.update({'math': __import__('math')})
        if use_python_cmath:
            namespace.update({'cmath': __import__('cmath')})
        if use_np:
            try:
                namespace.update({'np': __import__('numpy')})
            except ImportError:
                raise ImportError(
                    'experimental_lambdify failed to import numpy.')
        if use_interval:
            namespace.update({'imath': __import__(
                'sympy.plotting.intervalmath', fromlist=['intervalmath'])})
            namespace.update({'math': __import__('math')})

        # Construct the lambda
        if self.print_lambda:
            print(newexpr)
        eval_str = 'lambda %s : ( %s )' % (argstr, newexpr)
        self.eval_str = eval_str
        exec("MYNEWLAMBDA = %s" % eval_str, namespace)
        self.lambda_func = namespace['MYNEWLAMBDA']

    def __call__(self, *args, **kwargs):
        return self.lambda_func(*args, **kwargs)


    ##############################################################################
    # Dicts for translating from SymPy to other modules
    ##############################################################################
    ###
    # builtins
    ###
    # Functions with different names in builtins
    builtin_functions_different = {
        'Min': 'min',
        'Max': 'max',
        'Abs': 'abs',
    }

    # Strings that should be translated
    builtin_not_functions = {
        'I': '1j',
#        'oo': '1e400',
    }

    ###
    # numpy
    ###

    # Functions that are the same in numpy
    numpy_functions_same = [
        'sin', 'cos', 'tan', 'sinh', 'cosh', 'tanh', 'exp', 'log',
        'sqrt', 'floor', 'conjugate', 'sign',
    ]

    # Functions with different names in numpy
    numpy_functions_different = {
        "acos": "arccos",
        "acosh": "arccosh",
        "arg": "angle",
        "asin": "arcsin",
        "asinh": "arcsinh",
        "atan": "arctan",
        "atan2": "arctan2",
        "atanh": "arctanh",
        "ceiling": "ceil",
        "im": "imag",
        "ln": "log",
        "Max": "amax",
        "Min": "amin",
        "re": "real",
        "Abs": "abs",
    }

    # Strings that should be translated
    numpy_not_functions = {
        'pi': 'np.pi',
        'oo': 'np.inf',
        'E': 'np.e',
    }

    ###
    # Python math
    ###

    # Functions that are the same in math
    math_functions_same = [
        'sin', 'cos', 'tan', 'asin', 'acos', 'atan', 'atan2',
        'sinh', 'cosh', 'tanh', 'asinh', 'acosh', 'atanh',
        'exp', 'log', 'erf', 'sqrt', 'floor', 'factorial', 'gamma',
    ]

    # Functions with different names in math
    math_functions_different = {
        'ceiling': 'ceil',
        'ln': 'log',
        'loggamma': 'lgamma'
    }

    # Strings that should be translated
    math_not_functions = {
        'pi': 'math.pi',
        'E': 'math.e',
    }

    ###
    # Python cmath
    ###

    # Functions that are the same in cmath
    cmath_functions_same = [
        'sin', 'cos', 'tan', 'asin', 'acos', 'atan',
        'sinh', 'cosh', 'tanh', 'asinh', 'acosh', 'atanh',
        'exp', 'log', 'sqrt',
    ]

    # Functions with different names in cmath
    cmath_functions_different = {
        'ln': 'log',
        'arg': 'phase',
    }

    # Strings that should be translated
    cmath_not_functions = {
        'pi': 'cmath.pi',
        'E': 'cmath.e',
    }

    ###
    # intervalmath
    ###

    interval_not_functions = {
        'pi': 'math.pi',
        'E': 'math.e'
    }

    interval_functions_same = [
        'sin', 'cos', 'exp', 'tan', 'atan', 'log',
        'sqrt', 'cosh', 'sinh', 'tanh', 'floor',
        'acos', 'asin', 'acosh', 'asinh', 'atanh',
        'Abs', 'And', 'Or'
    ]

    interval_functions_different = {
        'Min': 'imin',
        'Max': 'imax',
        'ceiling': 'ceil',

    }

    ###
    # mpmath, etc
    ###
    #TODO

    ###
    # Create the final ordered tuples of dictionaries
    ###

    # For strings
    def get_dict_str(self):
        dict_str = dict(self.builtin_not_functions)
        if self.use_np:
            dict_str.update(self.numpy_not_functions)
        if self.use_python_math:
            dict_str.update(self.math_not_functions)
        if self.use_python_cmath:
            dict_str.update(self.cmath_not_functions)
        if self.use_interval:
            dict_str.update(self.interval_not_functions)
        return dict_str

    # For functions
    def get_dict_fun(self):
        dict_fun = dict(self.builtin_functions_different)
        if self.use_np:
            for s in self.numpy_functions_same:
                dict_fun[s] = 'np.' + s
            for k, v in self.numpy_functions_different.items():
                dict_fun[k] = 'np.' + v
        if self.use_python_math:
            for s in self.math_functions_same:
                dict_fun[s] = 'math.' + s
            for k, v in self.math_functions_different.items():
                dict_fun[k] = 'math.' + v
        if self.use_python_cmath:
            for s in self.cmath_functions_same:
                dict_fun[s] = 'cmath.' + s
            for k, v in self.cmath_functions_different.items():
                dict_fun[k] = 'cmath.' + v
        if self.use_interval:
            for s in self.interval_functions_same:
                dict_fun[s] = 'imath.' + s
            for k, v in self.interval_functions_different.items():
                dict_fun[k] = 'imath.' + v
        return dict_fun

    ##############################################################################
    # The translator functions, tree parsers, etc.
    ##############################################################################

    def str2tree(self, exprstr):
        """Converts an expression string to a tree.

        Explanation
        ===========

        Functions are represented by ('func_name(', tree_of_arguments).
        Other expressions are (head_string, mid_tree, tail_str).
        Expressions that do not contain functions are directly returned.

        Examples
        ========

        >>> from sympy.abc import x, y, z
        >>> from sympy import Integral, sin
        >>> from sympy.plotting.experimental_lambdify import Lambdifier
        >>> str2tree = Lambdifier([x], x).str2tree

        >>> str2tree(str(Integral(x, (x, 1, y))))
        ('', ('Integral(', 'x, (x, 1, y)'), ')')
        >>> str2tree(str(x+y))
        'x + y'
        >>> str2tree(str(x+y*sin(z)+1))
        ('x + y*', ('sin(', 'z'), ') + 1')
        >>> str2tree('sin(y*(y + 1.1) + (sin(y)))')
        ('', ('sin(', ('y*(y + 1.1) + (', ('sin(', 'y'), '))')), ')')
        """
        #matches the first 'function_name('
        first_par = re.search(r'(\w+\()', exprstr)
        if first_par is None:
            return exprstr
        else:
            start = first_par.start()
            end = first_par.end()
            head = exprstr[:start]
            func = exprstr[start:end]
            tail = exprstr[end:]
            count = 0
            for i, c in enumerate(tail):
                if c == '(':
                    count += 1
                elif c == ')':
                    count -= 1
                if count == -1:
                    break
            func_tail = self.str2tree(tail[:i])
            tail = self.str2tree(tail[i:])
            return (head, (func, func_tail), tail)

    @classmethod
    def tree2str(cls, tree):
        """Converts a tree to string without translations.

        Examples
        ========

        >>> from sympy.abc import x, y, z
        >>> from sympy import sin
        >>> from sympy.plotting.experimental_lambdify import Lambdifier
        >>> str2tree = Lambdifier([x], x).str2tree
        >>> tree2str = Lambdifier([x], x).tree2str

        >>> tree2str(str2tree(str(x+y*sin(z)+1)))
        'x + y*sin(z) + 1'
        """
        if isinstance(tree, str):
            return tree
        else:
            return ''.join(map(cls.tree2str, tree))

    def tree2str_translate(self, tree):
        """Converts a tree to string with translations.

        Explanation
        ===========

        Function names are translated by translate_func.
        Other strings are translated by translate_str.
        """
        if isinstance(tree, str):
            return self.translate_str(tree)
        elif isinstance(tree, tuple) and len(tree) == 2:
            return self.translate_func(tree[0][:-1], tree[1])
        else:
            return ''.join([self.tree2str_translate(t) for t in tree])

    def translate_str(self, estr):
        """Translate substrings of estr using in order the dictionaries in
        dict_tuple_str."""
        for pattern, repl in self.dict_str.items():
                estr = re.sub(pattern, repl, estr)
        return estr

    def translate_func(self, func_name, argtree):
        """Translate function names and the tree of arguments.

        Explanation
        ===========

        If the function name is not in the dictionaries of dict_tuple_fun then the
        function is surrounded by a float((...).evalf()).

        The use of float is necessary as np.<function>(sympy.Float(..)) raises an
        error."""
        if func_name in self.dict_fun:
            new_name = self.dict_fun[func_name]
            argstr = self.tree2str_translate(argtree)
            return new_name + '(' + argstr
        elif func_name in ['Eq', 'Ne']:
            op = {'Eq': '==', 'Ne': '!='}
            return "(lambda x, y: x {} y)({}".format(op[func_name], self.tree2str_translate(argtree))
        else:
            template = '(%s(%s)).evalf(' if self.use_evalf else '%s(%s'
            if self.float_wrap_evalf:
                template = 'float(%s)' % template
            elif self.complex_wrap_evalf:
                template = 'complex(%s)' % template

            # Wrapping should only happen on the outermost expression, which
            # is the only thing we know will be a number.
            float_wrap_evalf = self.float_wrap_evalf
            complex_wrap_evalf = self.complex_wrap_evalf
            self.float_wrap_evalf = False
            self.complex_wrap_evalf = False
            ret =  template % (func_name, self.tree2str_translate(argtree))
            self.float_wrap_evalf = float_wrap_evalf
            self.complex_wrap_evalf = complex_wrap_evalf
            return ret

    ##############################################################################
    # The namespace constructors
    ##############################################################################

    @classmethod
    def sympy_expression_namespace(cls, expr):
        """Traverses the (func, args) tree of an expression and creates a SymPy
        namespace. All other modules are imported only as a module name. That way
        the namespace is not polluted and rests quite small. It probably causes much
        more variable lookups and so it takes more time, but there are no tests on
        that for the moment."""
        if expr is None:
            return {}
        else:
            funcname = str(expr.func)
            # XXX Workaround
            # Here we add an ugly workaround because str(func(x))
            # is not always the same as str(func). Eg
            # >>> str(Integral(x))
            # "Integral(x)"
            # >>> str(Integral)
            # "<class 'sympy.integrals.integrals.Integral'>"
            # >>> str(sqrt(x))
            # "sqrt(x)"
            # >>> str(sqrt)
            # "<function sqrt at 0x3d92de8>"
            # >>> str(sin(x))
            # "sin(x)"
            # >>> str(sin)
            # "sin"
            # Either one of those can be used but not all at the same time.
            # The code considers the sin example as the right one.
            regexlist = [
                r'<class \'sympy[\w.]*?.([\w]*)\'>$',
                # the example Integral
                r'<function ([\w]*) at 0x[\w]*>$',    # the example sqrt
            ]
            for r in regexlist:
                m = re.match(r, funcname)
                if m is not None:
                    funcname = m.groups()[0]
            # End of the workaround
            # XXX debug: print funcname
            args_dict = {}
            for a in expr.args:
                if (isinstance(a, (Symbol, NumberSymbol)) or a in [I, zoo, oo]):
                    continue
                else:
                    args_dict.update(cls.sympy_expression_namespace(a))
            args_dict.update({funcname: expr.func})
            return args_dict

    @staticmethod
    def sympy_atoms_namespace(expr):
        """For no real reason this function is separated from
        sympy_expression_namespace. It can be moved to it."""
        atoms = expr.atoms(Symbol, NumberSymbol, I, zoo, oo)
        d = {}
        for a in atoms:
            # XXX debug: print 'atom:' + str(a)
            d[str(a)] = a
        return d

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\io\excel\_openpyxl.py ===
from __future__ import annotations

import mmap
from typing import (
    TYPE_CHECKING,
    Any,
    cast,
)

import numpy as np

from pandas.compat._optional import import_optional_dependency
from pandas.util._decorators import doc

from pandas.core.shared_docs import _shared_docs

from pandas.io.excel._base import (
    BaseExcelReader,
    ExcelWriter,
)
from pandas.io.excel._util import (
    combine_kwargs,
    validate_freeze_panes,
)

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.descriptors.serialisable import Serialisable

    from pandas._typing import (
        ExcelWriterIfSheetExists,
        FilePath,
        ReadBuffer,
        Scalar,
        StorageOptions,
        WriteExcelBuffer,
    )


class OpenpyxlWriter(ExcelWriter):
    _engine = "openpyxl"
    _supported_extensions = (".xlsx", ".xlsm")

    def __init__(
        self,
        path: FilePath | WriteExcelBuffer | ExcelWriter,
        engine: str | None = None,
        date_format: str | None = None,
        datetime_format: str | None = None,
        mode: str = "w",
        storage_options: StorageOptions | None = None,
        if_sheet_exists: ExcelWriterIfSheetExists | None = None,
        engine_kwargs: dict[str, Any] | None = None,
        **kwargs,
    ) -> None:
        # Use the openpyxl module as the Excel writer.
        from openpyxl.workbook import Workbook

        engine_kwargs = combine_kwargs(engine_kwargs, kwargs)

        super().__init__(
            path,
            mode=mode,
            storage_options=storage_options,
            if_sheet_exists=if_sheet_exists,
            engine_kwargs=engine_kwargs,
        )

        # ExcelWriter replaced "a" by "r+" to allow us to first read the excel file from
        # the file and later write to it
        if "r+" in self._mode:  # Load from existing workbook
            from openpyxl import load_workbook

            try:
                self._book = load_workbook(self._handles.handle, **engine_kwargs)
            except TypeError:
                self._handles.handle.close()
                raise
            self._handles.handle.seek(0)
        else:
            # Create workbook object with default optimized_write=True.
            try:
                self._book = Workbook(**engine_kwargs)
            except TypeError:
                self._handles.handle.close()
                raise

            if self.book.worksheets:
                self.book.remove(self.book.worksheets[0])

    @property
    def book(self) -> Workbook:
        """
        Book instance of class openpyxl.workbook.Workbook.

        This attribute can be used to access engine-specific features.
        """
        return self._book

    @property
    def sheets(self) -> dict[str, Any]:
        """Mapping of sheet names to sheet objects."""
        result = {name: self.book[name] for name in self.book.sheetnames}
        return result

    def _save(self) -> None:
        """
        Save workbook to disk.
        """
        self.book.save(self._handles.handle)
        if "r+" in self._mode and not isinstance(self._handles.handle, mmap.mmap):
            # truncate file to the written content
            self._handles.handle.truncate()

    @classmethod
    def _convert_to_style_kwargs(cls, style_dict: dict) -> dict[str, Serialisable]:
        """
        Convert a style_dict to a set of kwargs suitable for initializing
        or updating-on-copy an openpyxl v2 style object.

        Parameters
        ----------
        style_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'font'
                'fill'
                'border' ('borders')
                'alignment'
                'number_format'
                'protection'

        Returns
        -------
        style_kwargs : dict
            A dict with the same, normalized keys as ``style_dict`` but each
            value has been replaced with a native openpyxl style object of the
            appropriate class.
        """
        _style_key_map = {"borders": "border"}

        style_kwargs: dict[str, Serialisable] = {}
        for k, v in style_dict.items():
            k = _style_key_map.get(k, k)
            _conv_to_x = getattr(cls, f"_convert_to_{k}", lambda x: None)
            new_v = _conv_to_x(v)
            if new_v:
                style_kwargs[k] = new_v

        return style_kwargs

    @classmethod
    def _convert_to_color(cls, color_spec):
        """
        Convert ``color_spec`` to an openpyxl v2 Color object.

        Parameters
        ----------
        color_spec : str, dict
            A 32-bit ARGB hex string, or a dict with zero or more of the
            following keys.
                'rgb'
                'indexed'
                'auto'
                'theme'
                'tint'
                'index'
                'type'

        Returns
        -------
        color : openpyxl.styles.Color
        """
        from openpyxl.styles import Color

        if isinstance(color_spec, str):
            return Color(color_spec)
        else:
            return Color(**color_spec)

    @classmethod
    def _convert_to_font(cls, font_dict):
        """
        Convert ``font_dict`` to an openpyxl v2 Font object.

        Parameters
        ----------
        font_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'name'
                'size' ('sz')
                'bold' ('b')
                'italic' ('i')
                'underline' ('u')
                'strikethrough' ('strike')
                'color'
                'vertAlign' ('vertalign')
                'charset'
                'scheme'
                'family'
                'outline'
                'shadow'
                'condense'

        Returns
        -------
        font : openpyxl.styles.Font
        """
        from openpyxl.styles import Font

        _font_key_map = {
            "sz": "size",
            "b": "bold",
            "i": "italic",
            "u": "underline",
            "strike": "strikethrough",
            "vertalign": "vertAlign",
        }

        font_kwargs = {}
        for k, v in font_dict.items():
            k = _font_key_map.get(k, k)
            if k == "color":
                v = cls._convert_to_color(v)
            font_kwargs[k] = v

        return Font(**font_kwargs)

    @classmethod
    def _convert_to_stop(cls, stop_seq):
        """
        Convert ``stop_seq`` to a list of openpyxl v2 Color objects,
        suitable for initializing the ``GradientFill`` ``stop`` parameter.

        Parameters
        ----------
        stop_seq : iterable
            An iterable that yields objects suitable for consumption by
            ``_convert_to_color``.

        Returns
        -------
        stop : list of openpyxl.styles.Color
        """
        return map(cls._convert_to_color, stop_seq)

    @classmethod
    def _convert_to_fill(cls, fill_dict: dict[str, Any]):
        """
        Convert ``fill_dict`` to an openpyxl v2 Fill object.

        Parameters
        ----------
        fill_dict : dict
            A dict with one or more of the following keys (or their synonyms),
                'fill_type' ('patternType', 'patterntype')
                'start_color' ('fgColor', 'fgcolor')
                'end_color' ('bgColor', 'bgcolor')
            or one or more of the following keys (or their synonyms).
                'type' ('fill_type')
                'degree'
                'left'
                'right'
                'top'
                'bottom'
                'stop'

        Returns
        -------
        fill : openpyxl.styles.Fill
        """
        from openpyxl.styles import (
            GradientFill,
            PatternFill,
        )

        _pattern_fill_key_map = {
            "patternType": "fill_type",
            "patterntype": "fill_type",
            "fgColor": "start_color",
            "fgcolor": "start_color",
            "bgColor": "end_color",
            "bgcolor": "end_color",
        }

        _gradient_fill_key_map = {"fill_type": "type"}

        pfill_kwargs = {}
        gfill_kwargs = {}
        for k, v in fill_dict.items():
            pk = _pattern_fill_key_map.get(k)
            gk = _gradient_fill_key_map.get(k)
            if pk in ["start_color", "end_color"]:
                v = cls._convert_to_color(v)
            if gk == "stop":
                v = cls._convert_to_stop(v)
            if pk:
                pfill_kwargs[pk] = v
            elif gk:
                gfill_kwargs[gk] = v
            else:
                pfill_kwargs[k] = v
                gfill_kwargs[k] = v

        try:
            return PatternFill(**pfill_kwargs)
        except TypeError:
            return GradientFill(**gfill_kwargs)

    @classmethod
    def _convert_to_side(cls, side_spec):
        """
        Convert ``side_spec`` to an openpyxl v2 Side object.

        Parameters
        ----------
        side_spec : str, dict
            A string specifying the border style, or a dict with zero or more
            of the following keys (or their synonyms).
                'style' ('border_style')
                'color'

        Returns
        -------
        side : openpyxl.styles.Side
        """
        from openpyxl.styles import Side

        _side_key_map = {"border_style": "style"}

        if isinstance(side_spec, str):
            return Side(style=side_spec)

        side_kwargs = {}
        for k, v in side_spec.items():
            k = _side_key_map.get(k, k)
            if k == "color":
                v = cls._convert_to_color(v)
            side_kwargs[k] = v

        return Side(**side_kwargs)

    @classmethod
    def _convert_to_border(cls, border_dict):
        """
        Convert ``border_dict`` to an openpyxl v2 Border object.

        Parameters
        ----------
        border_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'left'
                'right'
                'top'
                'bottom'
                'diagonal'
                'diagonal_direction'
                'vertical'
                'horizontal'
                'diagonalUp' ('diagonalup')
                'diagonalDown' ('diagonaldown')
                'outline'

        Returns
        -------
        border : openpyxl.styles.Border
        """
        from openpyxl.styles import Border

        _border_key_map = {"diagonalup": "diagonalUp", "diagonaldown": "diagonalDown"}

        border_kwargs = {}
        for k, v in border_dict.items():
            k = _border_key_map.get(k, k)
            if k == "color":
                v = cls._convert_to_color(v)
            if k in ["left", "right", "top", "bottom", "diagonal"]:
                v = cls._convert_to_side(v)
            border_kwargs[k] = v

        return Border(**border_kwargs)

    @classmethod
    def _convert_to_alignment(cls, alignment_dict):
        """
        Convert ``alignment_dict`` to an openpyxl v2 Alignment object.

        Parameters
        ----------
        alignment_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'horizontal'
                'vertical'
                'text_rotation'
                'wrap_text'
                'shrink_to_fit'
                'indent'
        Returns
        -------
        alignment : openpyxl.styles.Alignment
        """
        from openpyxl.styles import Alignment

        return Alignment(**alignment_dict)

    @classmethod
    def _convert_to_number_format(cls, number_format_dict):
        """
        Convert ``number_format_dict`` to an openpyxl v2.1.0 number format
        initializer.

        Parameters
        ----------
        number_format_dict : dict
            A dict with zero or more of the following keys.
                'format_code' : str

        Returns
        -------
        number_format : str
        """
        return number_format_dict["format_code"]

    @classmethod
    def _convert_to_protection(cls, protection_dict):
        """
        Convert ``protection_dict`` to an openpyxl v2 Protection object.

        Parameters
        ----------
        protection_dict : dict
            A dict with zero or more of the following keys.
                'locked'
                'hidden'

        Returns
        -------
        """
        from openpyxl.styles import Protection

        return Protection(**protection_dict)

    def _write_cells(
        self,
        cells,
        sheet_name: str | None = None,
        startrow: int = 0,
        startcol: int = 0,
        freeze_panes: tuple[int, int] | None = None,
    ) -> None:
        # Write the frame cells using openpyxl.
        sheet_name = self._get_sheet_name(sheet_name)

        _style_cache: dict[str, dict[str, Serialisable]] = {}

        if sheet_name in self.sheets and self._if_sheet_exists != "new":
            if "r+" in self._mode:
                if self._if_sheet_exists == "replace":
                    old_wks = self.sheets[sheet_name]
                    target_index = self.book.index(old_wks)
                    del self.book[sheet_name]
                    wks = self.book.create_sheet(sheet_name, target_index)
                elif self._if_sheet_exists == "error":
                    raise ValueError(
                        f"Sheet '{sheet_name}' already exists and "
                        f"if_sheet_exists is set to 'error'."
                    )
                elif self._if_sheet_exists == "overlay":
                    wks = self.sheets[sheet_name]
                else:
                    raise ValueError(
                        f"'{self._if_sheet_exists}' is not valid for if_sheet_exists. "
                        "Valid options are 'error', 'new', 'replace' and 'overlay'."
                    )
            else:
                wks = self.sheets[sheet_name]
        else:
            wks = self.book.create_sheet()
            wks.title = sheet_name

        if validate_freeze_panes(freeze_panes):
            freeze_panes = cast(tuple[int, int], freeze_panes)
            wks.freeze_panes = wks.cell(
                row=freeze_panes[0] + 1, column=freeze_panes[1] + 1
            )

        for cell in cells:
            xcell = wks.cell(
                row=startrow + cell.row + 1, column=startcol + cell.col + 1
            )
            xcell.value, fmt = self._value_with_fmt(cell.val)
            if fmt:
                xcell.number_format = fmt

            style_kwargs: dict[str, Serialisable] | None = {}
            if cell.style:
                key = str(cell.style)
                style_kwargs = _style_cache.get(key)
                if style_kwargs is None:
                    style_kwargs = self._convert_to_style_kwargs(cell.style)
                    _style_cache[key] = style_kwargs

            if style_kwargs:
                for k, v in style_kwargs.items():
                    setattr(xcell, k, v)

            if cell.mergestart is not None and cell.mergeend is not None:
                wks.merge_cells(
                    start_row=startrow + cell.row + 1,
                    start_column=startcol + cell.col + 1,
                    end_column=startcol + cell.mergeend + 1,
                    end_row=startrow + cell.mergestart + 1,
                )

                # When cells are merged only the top-left cell is preserved
                # The behaviour of the other cells in a merged range is
                # undefined
                if style_kwargs:
                    first_row = startrow + cell.row + 1
                    last_row = startrow + cell.mergestart + 1
                    first_col = startcol + cell.col + 1
                    last_col = startcol + cell.mergeend + 1

                    for row in range(first_row, last_row + 1):
                        for col in range(first_col, last_col + 1):
                            if row == first_row and col == first_col:
                                # Ignore first cell. It is already handled.
                                continue
                            xcell = wks.cell(column=col, row=row)
                            for k, v in style_kwargs.items():
                                setattr(xcell, k, v)


class OpenpyxlReader(BaseExcelReader["Workbook"]):
    @doc(storage_options=_shared_docs["storage_options"])
    def __init__(
        self,
        filepath_or_buffer: FilePath | ReadBuffer[bytes],
        storage_options: StorageOptions | None = None,
        engine_kwargs: dict | None = None,
    ) -> None:
        """
        Reader using openpyxl engine.

        Parameters
        ----------
        filepath_or_buffer : str, path object or Workbook
            Object to be parsed.
        {storage_options}
        engine_kwargs : dict, optional
            Arbitrary keyword arguments passed to excel engine.
        """
        import_optional_dependency("openpyxl")
        super().__init__(
            filepath_or_buffer,
            storage_options=storage_options,
            engine_kwargs=engine_kwargs,
        )

    @property
    def _workbook_class(self) -> type[Workbook]:
        from openpyxl import Workbook

        return Workbook

    def load_workbook(
        self, filepath_or_buffer: FilePath | ReadBuffer[bytes], engine_kwargs
    ) -> Workbook:
        from openpyxl import load_workbook

        default_kwargs = {"read_only": True, "data_only": True, "keep_links": False}

        return load_workbook(
            filepath_or_buffer,
            **(default_kwargs | engine_kwargs),
        )

    @property
    def sheet_names(self) -> list[str]:
        return [sheet.title for sheet in self.book.worksheets]

    def get_sheet_by_name(self, name: str):
        self.raise_if_bad_sheet_by_name(name)
        return self.book[name]

    def get_sheet_by_index(self, index: int):
        self.raise_if_bad_sheet_by_index(index)
        return self.book.worksheets[index]

    def _convert_cell(self, cell) -> Scalar:
        from openpyxl.cell.cell import (
            TYPE_ERROR,
            TYPE_NUMERIC,
        )

        if cell.value is None:
            return ""  # compat with xlrd
        elif cell.data_type == TYPE_ERROR:
            return np.nan
        elif cell.data_type == TYPE_NUMERIC:
            val = int(cell.value)
            if val == cell.value:
                return val
            return float(cell.value)

        return cell.value

    def get_sheet_data(
        self, sheet, file_rows_needed: int | None = None
    ) -> list[list[Scalar]]:
        if self.book.read_only:
            sheet.reset_dimensions()

        data: list[list[Scalar]] = []
        last_row_with_data = -1
        for row_number, row in enumerate(sheet.rows):
            converted_row = [self._convert_cell(cell) for cell in row]
            while converted_row and converted_row[-1] == "":
                # trim trailing empty elements
                converted_row.pop()
            if converted_row:
                last_row_with_data = row_number
            data.append(converted_row)
            if file_rows_needed is not None and len(data) >= file_rows_needed:
                break

        # Trim trailing empty rows
        data = data[: last_row_with_data + 1]

        if len(data) > 0:
            # extend rows to max width
            max_width = max(len(data_row) for data_row in data)
            if min(len(data_row) for data_row in data) < max_width:
                empty_cell: list[Scalar] = [""]
                data = [
                    data_row + (max_width - len(data_row)) * empty_cell
                    for data_row in data
                ]

        return data

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\fields.py ===
"""Sparse rational function fields. """

from __future__ import annotations
from functools import reduce

from operator import add, mul, lt, le, gt, ge

from sympy.core.expr import Expr
from sympy.core.mod import Mod
from sympy.core.numbers import Exp1
from sympy.core.singleton import S
from sympy.core.symbol import Symbol
from sympy.core.sympify import CantSympify, sympify
from sympy.functions.elementary.exponential import ExpBase
from sympy.polys.domains.domain import Domain
from sympy.polys.domains.domainelement import DomainElement
from sympy.polys.domains.fractionfield import FractionField
from sympy.polys.domains.polynomialring import PolynomialRing
from sympy.polys.constructor import construct_domain
from sympy.polys.orderings import lex, MonomialOrder
from sympy.polys.polyerrors import CoercionFailed
from sympy.polys.polyoptions import build_options
from sympy.polys.polyutils import _parallel_dict_from_expr
from sympy.polys.rings import PolyRing, PolyElement
from sympy.printing.defaults import DefaultPrinting
from sympy.utilities import public
from sympy.utilities.iterables import is_sequence
from sympy.utilities.magic import pollute

@public
def field(symbols, domain, order=lex):
    """Construct new rational function field returning (field, x1, ..., xn). """
    _field = FracField(symbols, domain, order)
    return (_field,) + _field.gens

@public
def xfield(symbols, domain, order=lex):
    """Construct new rational function field returning (field, (x1, ..., xn)). """
    _field = FracField(symbols, domain, order)
    return (_field, _field.gens)

@public
def vfield(symbols, domain, order=lex):
    """Construct new rational function field and inject generators into global namespace. """
    _field = FracField(symbols, domain, order)
    pollute([ sym.name for sym in _field.symbols ], _field.gens)
    return _field

@public
def sfield(exprs, *symbols, **options):
    """Construct a field deriving generators and domain
    from options and input expressions.

    Parameters
    ==========

    exprs   : py:class:`~.Expr` or sequence of :py:class:`~.Expr` (sympifiable)

    symbols : sequence of :py:class:`~.Symbol`/:py:class:`~.Expr`

    options : keyword arguments understood by :py:class:`~.Options`

    Examples
    ========

    >>> from sympy import exp, log, symbols, sfield

    >>> x = symbols("x")
    >>> K, f = sfield((x*log(x) + 4*x**2)*exp(1/x + log(x)/3)/x**2)
    >>> K
    Rational function field in x, exp(1/x), log(x), x**(1/3) over ZZ with lex order
    >>> f
    (4*x**2*(exp(1/x)) + x*(exp(1/x))*(log(x)))/((x**(1/3))**5)
    """
    single = False
    if not is_sequence(exprs):
        exprs, single = [exprs], True

    exprs = list(map(sympify, exprs))
    opt = build_options(symbols, options)
    numdens = []
    for expr in exprs:
        numdens.extend(expr.as_numer_denom())
    reps, opt = _parallel_dict_from_expr(numdens, opt)

    if opt.domain is None:
        # NOTE: this is inefficient because construct_domain() automatically
        # performs conversion to the target domain. It shouldn't do this.
        coeffs = sum([list(rep.values()) for rep in reps], [])
        opt.domain, _ = construct_domain(coeffs, opt=opt)

    _field = FracField(opt.gens, opt.domain, opt.order)
    fracs = []
    for i in range(0, len(reps), 2):
        fracs.append(_field(tuple(reps[i:i+2])))

    if single:
        return (_field, fracs[0])
    else:
        return (_field, fracs)


class FracField(DefaultPrinting):
    """Multivariate distributed rational function field. """

    ring: PolyRing
    gens: tuple[FracElement, ...]
    symbols: tuple[Expr, ...]
    ngens: int
    domain: Domain
    order: MonomialOrder

    def __new__(cls, symbols, domain, order=lex):
        ring = PolyRing(symbols, domain, order)
        symbols = ring.symbols
        ngens = ring.ngens
        domain = ring.domain
        order = ring.order

        _hash_tuple = (cls.__name__, symbols, ngens, domain, order)

        obj = object.__new__(cls)
        obj._hash_tuple = _hash_tuple
        obj._hash = hash(_hash_tuple)
        obj.ring = ring
        obj.symbols = symbols
        obj.ngens = ngens
        obj.domain = domain
        obj.order = order

        obj.dtype = FracElement(obj, ring.zero).raw_new

        obj.zero = obj.dtype(ring.zero)
        obj.one = obj.dtype(ring.one)

        obj.gens = obj._gens()

        for symbol, generator in zip(obj.symbols, obj.gens):
            if isinstance(symbol, Symbol):
                name = symbol.name

                if not hasattr(obj, name):
                    setattr(obj, name, generator)

        return obj

    def _gens(self):
        """Return a list of polynomial generators. """
        return tuple([ self.dtype(gen) for gen in self.ring.gens ])

    def __getnewargs__(self):
        return (self.symbols, self.domain, self.order)

    def __hash__(self):
        return self._hash

    def index(self, gen):
        if self.is_element(gen):
            return self.ring.index(gen.to_poly())
        else:
            raise ValueError("expected a %s, got %s instead" % (self.dtype,gen))

    def __eq__(self, other):
        return isinstance(other, FracField) and \
            (self.symbols, self.ngens, self.domain, self.order) == \
            (other.symbols, other.ngens, other.domain, other.order)

    def __ne__(self, other):
        return not self == other

    def is_element(self, element):
        """True if ``element`` is an element of this field. False otherwise. """
        return isinstance(element, FracElement) and element.field == self

    def raw_new(self, numer, denom=None):
        return self.dtype(numer, denom)

    def new(self, numer, denom=None):
        if denom is None: denom = self.ring.one
        numer, denom = numer.cancel(denom)
        return self.raw_new(numer, denom)

    def domain_new(self, element):
        return self.domain.convert(element)

    def ground_new(self, element):
        try:
            return self.new(self.ring.ground_new(element))
        except CoercionFailed:
            domain = self.domain

            if not domain.is_Field and domain.has_assoc_Field:
                ring = self.ring
                ground_field = domain.get_field()
                element = ground_field.convert(element)
                numer = ring.ground_new(ground_field.numer(element))
                denom = ring.ground_new(ground_field.denom(element))
                return self.raw_new(numer, denom)
            else:
                raise

    def field_new(self, element):
        if isinstance(element, FracElement):
            if self == element.field:
                return element

            if isinstance(self.domain, FractionField) and \
                self.domain.field == element.field:
                return self.ground_new(element)
            elif isinstance(self.domain, PolynomialRing) and \
                self.domain.ring.to_field() == element.field:
                return self.ground_new(element)
            else:
                raise NotImplementedError("conversion")
        elif isinstance(element, PolyElement):
            denom, numer = element.clear_denoms()

            if isinstance(self.domain, PolynomialRing) and \
                numer.ring == self.domain.ring:
                numer = self.ring.ground_new(numer)
            elif isinstance(self.domain, FractionField) and \
                numer.ring == self.domain.field.to_ring():
                numer = self.ring.ground_new(numer)
            else:
                numer = numer.set_ring(self.ring)

            denom = self.ring.ground_new(denom)
            return self.raw_new(numer, denom)
        elif isinstance(element, tuple) and len(element) == 2:
            numer, denom = list(map(self.ring.ring_new, element))
            return self.new(numer, denom)
        elif isinstance(element, str):
            raise NotImplementedError("parsing")
        elif isinstance(element, Expr):
            return self.from_expr(element)
        else:
            return self.ground_new(element)

    __call__ = field_new

    def _rebuild_expr(self, expr, mapping):
        domain = self.domain
        powers = tuple((gen, gen.as_base_exp()) for gen in mapping.keys()
            if gen.is_Pow or isinstance(gen, ExpBase))

        def _rebuild(expr):
            generator = mapping.get(expr)

            if generator is not None:
                return generator
            elif expr.is_Add:
                return reduce(add, list(map(_rebuild, expr.args)))
            elif expr.is_Mul:
                return reduce(mul, list(map(_rebuild, expr.args)))
            elif expr.is_Pow or isinstance(expr, (ExpBase, Exp1)):
                b, e = expr.as_base_exp()
                # look for bg**eg whose integer power may be b**e
                for gen, (bg, eg) in powers:
                    if bg == b and Mod(e, eg) == 0:
                        return mapping.get(gen)**int(e/eg)
                if e.is_Integer and e is not S.One:
                    return _rebuild(b)**int(e)
            elif mapping.get(1/expr) is not None:
                return 1/mapping.get(1/expr)

            try:
                return domain.convert(expr)
            except CoercionFailed:
                if not domain.is_Field and domain.has_assoc_Field:
                    return domain.get_field().convert(expr)
                else:
                    raise

        return _rebuild(expr)

    def from_expr(self, expr):
        mapping = dict(list(zip(self.symbols, self.gens)))

        try:
            frac = self._rebuild_expr(sympify(expr), mapping)
        except CoercionFailed:
            raise ValueError("expected an expression convertible to a rational function in %s, got %s" % (self, expr))
        else:
            return self.field_new(frac)

    def to_domain(self):
        return FractionField(self)

    def to_ring(self):
        return PolyRing(self.symbols, self.domain, self.order)

class FracElement(DomainElement, DefaultPrinting, CantSympify):
    """Element of multivariate distributed rational function field. """

    def __init__(self, field, numer, denom=None):
        if denom is None:
            denom = field.ring.one
        elif not denom:
            raise ZeroDivisionError("zero denominator")

        self.field = field
        self.numer = numer
        self.denom = denom

    def raw_new(f, numer, denom=None):
        return f.__class__(f.field, numer, denom)

    def new(f, numer, denom):
        return f.raw_new(*numer.cancel(denom))

    def to_poly(f):
        if f.denom != 1:
            raise ValueError("f.denom should be 1")
        return f.numer

    def parent(self):
        return self.field.to_domain()

    def __getnewargs__(self):
        return (self.field, self.numer, self.denom)

    _hash = None

    def __hash__(self):
        _hash = self._hash
        if _hash is None:
            self._hash = _hash = hash((self.field, self.numer, self.denom))
        return _hash

    def copy(self):
        return self.raw_new(self.numer.copy(), self.denom.copy())

    def set_field(self, new_field):
        if self.field == new_field:
            return self
        else:
            new_ring = new_field.ring
            numer = self.numer.set_ring(new_ring)
            denom = self.denom.set_ring(new_ring)
            return new_field.new(numer, denom)

    def as_expr(self, *symbols):
        return self.numer.as_expr(*symbols)/self.denom.as_expr(*symbols)

    def __eq__(f, g):
        if isinstance(g, FracElement) and f.field == g.field:
            return f.numer == g.numer and f.denom == g.denom
        else:
            return f.numer == g and f.denom == f.field.ring.one

    def __ne__(f, g):
        return not f == g

    def __bool__(f):
        return bool(f.numer)

    def sort_key(self):
        return (self.denom.sort_key(), self.numer.sort_key())

    def _cmp(f1, f2, op):
        if f1.field.is_element(f2):
            return op(f1.sort_key(), f2.sort_key())
        else:
            return NotImplemented

    def __lt__(f1, f2):
        return f1._cmp(f2, lt)
    def __le__(f1, f2):
        return f1._cmp(f2, le)
    def __gt__(f1, f2):
        return f1._cmp(f2, gt)
    def __ge__(f1, f2):
        return f1._cmp(f2, ge)

    def __pos__(f):
        """Negate all coefficients in ``f``. """
        return f.raw_new(f.numer, f.denom)

    def __neg__(f):
        """Negate all coefficients in ``f``. """
        return f.raw_new(-f.numer, f.denom)

    def _extract_ground(self, element):
        domain = self.field.domain

        try:
            element = domain.convert(element)
        except CoercionFailed:
            if not domain.is_Field and domain.has_assoc_Field:
                ground_field = domain.get_field()

                try:
                    element = ground_field.convert(element)
                except CoercionFailed:
                    pass
                else:
                    return -1, ground_field.numer(element), ground_field.denom(element)

            return 0, None, None
        else:
            return 1, element, None

    def __add__(f, g):
        """Add rational functions ``f`` and ``g``. """
        field = f.field

        if not g:
            return f
        elif not f:
            return g
        elif field.is_element(g):
            if f.denom == g.denom:
                return f.new(f.numer + g.numer, f.denom)
            else:
                return f.new(f.numer*g.denom + f.denom*g.numer, f.denom*g.denom)
        elif field.ring.is_element(g):
            return f.new(f.numer + f.denom*g, f.denom)
        else:
            if isinstance(g, FracElement):
                if isinstance(field.domain, FractionField) and field.domain.field == g.field:
                    pass
                elif isinstance(g.field.domain, FractionField) and g.field.domain.field == field:
                    return g.__radd__(f)
                else:
                    return NotImplemented
            elif isinstance(g, PolyElement):
                if isinstance(field.domain, PolynomialRing) and field.domain.ring == g.ring:
                    pass
                else:
                    return g.__radd__(f)

        return f.__radd__(g)

    def __radd__(f, c):
        if f.field.ring.is_element(c):
            return f.new(f.numer + f.denom*c, f.denom)

        op, g_numer, g_denom = f._extract_ground(c)

        if op == 1:
            return f.new(f.numer + f.denom*g_numer, f.denom)
        elif not op:
            return NotImplemented
        else:
            return f.new(f.numer*g_denom + f.denom*g_numer, f.denom*g_denom)

    def __sub__(f, g):
        """Subtract rational functions ``f`` and ``g``. """
        field = f.field

        if not g:
            return f
        elif not f:
            return -g
        elif field.is_element(g):
            if f.denom == g.denom:
                return f.new(f.numer - g.numer, f.denom)
            else:
                return f.new(f.numer*g.denom - f.denom*g.numer, f.denom*g.denom)
        elif field.ring.is_element(g):
            return f.new(f.numer - f.denom*g, f.denom)
        else:
            if isinstance(g, FracElement):
                if isinstance(field.domain, FractionField) and field.domain.field == g.field:
                    pass
                elif isinstance(g.field.domain, FractionField) and g.field.domain.field == field:
                    return g.__rsub__(f)
                else:
                    return NotImplemented
            elif isinstance(g, PolyElement):
                if isinstance(field.domain, PolynomialRing) and field.domain.ring == g.ring:
                    pass
                else:
                    return g.__rsub__(f)

        op, g_numer, g_denom = f._extract_ground(g)

        if op == 1:
            return f.new(f.numer - f.denom*g_numer, f.denom)
        elif not op:
            return NotImplemented
        else:
            return f.new(f.numer*g_denom - f.denom*g_numer, f.denom*g_denom)

    def __rsub__(f, c):
        if f.field.ring.is_element(c):
            return f.new(-f.numer + f.denom*c, f.denom)

        op, g_numer, g_denom = f._extract_ground(c)

        if op == 1:
            return f.new(-f.numer + f.denom*g_numer, f.denom)
        elif not op:
            return NotImplemented
        else:
            return f.new(-f.numer*g_denom + f.denom*g_numer, f.denom*g_denom)

    def __mul__(f, g):
        """Multiply rational functions ``f`` and ``g``. """
        field = f.field

        if not f or not g:
            return field.zero
        elif field.is_element(g):
            return f.new(f.numer*g.numer, f.denom*g.denom)
        elif field.ring.is_element(g):
            return f.new(f.numer*g, f.denom)
        else:
            if isinstance(g, FracElement):
                if isinstance(field.domain, FractionField) and field.domain.field == g.field:
                    pass
                elif isinstance(g.field.domain, FractionField) and g.field.domain.field == field:
                    return g.__rmul__(f)
                else:
                    return NotImplemented
            elif isinstance(g, PolyElement):
                if isinstance(field.domain, PolynomialRing) and field.domain.ring == g.ring:
                    pass
                else:
                    return g.__rmul__(f)

        return f.__rmul__(g)

    def __rmul__(f, c):
        if f.field.ring.is_element(c):
            return f.new(f.numer*c, f.denom)

        op, g_numer, g_denom = f._extract_ground(c)

        if op == 1:
            return f.new(f.numer*g_numer, f.denom)
        elif not op:
            return NotImplemented
        else:
            return f.new(f.numer*g_numer, f.denom*g_denom)

    def __truediv__(f, g):
        """Computes quotient of fractions ``f`` and ``g``. """
        field = f.field

        if not g:
            raise ZeroDivisionError
        elif field.is_element(g):
            return f.new(f.numer*g.denom, f.denom*g.numer)
        elif field.ring.is_element(g):
            return f.new(f.numer, f.denom*g)
        else:
            if isinstance(g, FracElement):
                if isinstance(field.domain, FractionField) and field.domain.field == g.field:
                    pass
                elif isinstance(g.field.domain, FractionField) and g.field.domain.field == field:
                    return g.__rtruediv__(f)
                else:
                    return NotImplemented
            elif isinstance(g, PolyElement):
                if isinstance(field.domain, PolynomialRing) and field.domain.ring == g.ring:
                    pass
                else:
                    return g.__rtruediv__(f)

        op, g_numer, g_denom = f._extract_ground(g)

        if op == 1:
            return f.new(f.numer, f.denom*g_numer)
        elif not op:
            return NotImplemented
        else:
            return f.new(f.numer*g_denom, f.denom*g_numer)

    def __rtruediv__(f, c):
        if not f:
            raise ZeroDivisionError
        elif f.field.ring.is_element(c):
            return f.new(f.denom*c, f.numer)

        op, g_numer, g_denom = f._extract_ground(c)

        if op == 1:
            return f.new(f.denom*g_numer, f.numer)
        elif not op:
            return NotImplemented
        else:
            return f.new(f.denom*g_numer, f.numer*g_denom)

    def __pow__(f, n):
        """Raise ``f`` to a non-negative power ``n``. """
        if n >= 0:
            return f.raw_new(f.numer**n, f.denom**n)
        elif not f:
            raise ZeroDivisionError
        else:
            return f.raw_new(f.denom**-n, f.numer**-n)

    def diff(f, x):
        """Computes partial derivative in ``x``.

        Examples
        ========

        >>> from sympy.polys.fields import field
        >>> from sympy.polys.domains import ZZ

        >>> _, x, y, z = field("x,y,z", ZZ)
        >>> ((x**2 + y)/(z + 1)).diff(x)
        2*x/(z + 1)

        """
        x = x.to_poly()
        return f.new(f.numer.diff(x)*f.denom - f.numer*f.denom.diff(x), f.denom**2)

    def __call__(f, *values):
        if 0 < len(values) <= f.field.ngens:
            return f.evaluate(list(zip(f.field.gens, values)))
        else:
            raise ValueError("expected at least 1 and at most %s values, got %s" % (f.field.ngens, len(values)))

    def evaluate(f, x, a=None):
        if isinstance(x, list) and a is None:
            x = [ (X.to_poly(), a) for X, a in x ]
            numer, denom = f.numer.evaluate(x), f.denom.evaluate(x)
        else:
            x = x.to_poly()
            numer, denom = f.numer.evaluate(x, a), f.denom.evaluate(x, a)

        field = numer.ring.to_field()
        return field.new(numer, denom)

    def subs(f, x, a=None):
        if isinstance(x, list) and a is None:
            x = [ (X.to_poly(), a) for X, a in x ]
            numer, denom = f.numer.subs(x), f.denom.subs(x)
        else:
            x = x.to_poly()
            numer, denom = f.numer.subs(x, a), f.denom.subs(x, a)

        return f.new(numer, denom)

    def compose(f, x, a=None):
        raise NotImplementedError

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\sam2\benchmark_sam2.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

"""
Benchmark performance of SAM2 encoder with ORT or PyTorch. See benchmark_sam2.sh for usage.
"""

import argparse
import csv
import statistics
import time
from collections.abc import Mapping
from datetime import datetime

import torch
from image_decoder import SAM2ImageDecoder
from image_encoder import SAM2ImageEncoder
from sam2_utils import decoder_shape_dict, encoder_shape_dict, load_sam2_model

from onnxruntime import InferenceSession, SessionOptions, get_available_providers
from onnxruntime.transformers.io_binding_helper import CudaSession


class TestConfig:
    def __init__(
        self,
        model_type: str,
        onnx_path: str,
        sam2_dir: str,
        device: torch.device,
        component: str = "image_encoder",
        provider="CPUExecutionProvider",
        torch_compile_mode="max-autotune",
        batch_size: int = 1,
        height: int = 1024,
        width: int = 1024,
        num_labels: int = 1,
        num_points: int = 1,
        num_masks: int = 1,
        multi_mask_output: bool = False,
        use_tf32: bool = True,
        enable_cuda_graph: bool = False,
        dtype=torch.float32,
        prefer_nhwc: bool = False,
        warm_up: int = 5,
        enable_nvtx_profile: bool = False,
        enable_ort_profile: bool = False,
        enable_torch_profile: bool = False,
        repeats: int = 1000,
        verbose: bool = False,
    ):
        assert model_type in ["sam2_hiera_tiny", "sam2_hiera_small", "sam2_hiera_large", "sam2_hiera_base_plus"]
        assert height >= 160 and height <= 4096
        assert width >= 160 and width <= 4096

        self.model_type = model_type
        self.onnx_path = onnx_path
        self.sam2_dir = sam2_dir
        self.component = component
        self.provider = provider
        self.torch_compile_mode = torch_compile_mode
        self.batch_size = batch_size
        self.height = height
        self.width = width
        self.num_labels = num_labels
        self.num_points = num_points
        self.num_masks = num_masks
        self.multi_mask_output = multi_mask_output
        self.device = device
        self.use_tf32 = use_tf32
        self.enable_cuda_graph = enable_cuda_graph
        self.dtype = dtype
        self.prefer_nhwc = prefer_nhwc
        self.warm_up = warm_up
        self.enable_nvtx_profile = enable_nvtx_profile
        self.enable_ort_profile = enable_ort_profile
        self.enable_torch_profile = enable_torch_profile
        self.repeats = repeats
        self.verbose = verbose

        if self.component == "image_encoder":
            assert self.height == 1024 and self.width == 1024, "Only image size 1024x1024 is allowed for image encoder."

    def __repr__(self):
        return f"{vars(self)}"

    def shape_dict(self) -> Mapping[str, list[int]]:
        if self.component == "image_encoder":
            return encoder_shape_dict(self.batch_size, self.height, self.width)
        else:
            return decoder_shape_dict(self.height, self.width, self.num_labels, self.num_points, self.num_masks)

    def random_inputs(self) -> Mapping[str, torch.Tensor]:
        dtype = self.dtype
        if self.component == "image_encoder":
            return {"image": torch.randn(self.batch_size, 3, self.height, self.width, dtype=dtype, device=self.device)}
        else:
            return {
                "image_features_0": torch.rand(1, 32, 256, 256, dtype=dtype, device=self.device),
                "image_features_1": torch.rand(1, 64, 128, 128, dtype=dtype, device=self.device),
                "image_embeddings": torch.rand(1, 256, 64, 64, dtype=dtype, device=self.device),
                "point_coords": torch.randint(
                    0, 1024, (self.num_labels, self.num_points, 2), dtype=dtype, device=self.device
                ),
                "point_labels": torch.randint(
                    0, 1, (self.num_labels, self.num_points), dtype=torch.int32, device=self.device
                ),
                "input_masks": torch.zeros(self.num_labels, 1, 256, 256, dtype=dtype, device=self.device),
                "has_input_masks": torch.ones(self.num_labels, dtype=dtype, device=self.device),
                "original_image_size": torch.tensor([self.height, self.width], dtype=torch.int32, device=self.device),
            }


def create_ort_session(config: TestConfig, session_options=None) -> InferenceSession:
    if config.verbose:
        print(f"create session for {vars(config)}")

    if config.provider == "CUDAExecutionProvider":
        device_id = torch.cuda.current_device() if isinstance(config.device, str) else config.device.index
        provider_options = CudaSession.get_cuda_provider_options(device_id, config.enable_cuda_graph)
        provider_options["use_tf32"] = int(config.use_tf32)
        if config.prefer_nhwc:
            provider_options["prefer_nhwc"] = 1
        providers = [(config.provider, provider_options), "CPUExecutionProvider"]
    else:
        providers = ["CPUExecutionProvider"]

    ort_session = InferenceSession(config.onnx_path, session_options, providers=providers)
    return ort_session


def create_session(config: TestConfig, session_options=None) -> CudaSession:
    ort_session = create_ort_session(config, session_options)
    cuda_session = CudaSession(ort_session, config.device, config.enable_cuda_graph)
    cuda_session.allocate_buffers(config.shape_dict())
    return cuda_session


class OrtTestSession:
    """A wrapper of ORT session to test relevance and performance."""

    def __init__(self, config: TestConfig, session_options=None):
        self.ort_session = create_session(config, session_options)
        self.feed_dict = config.random_inputs()

    def infer(self):
        return self.ort_session.infer(self.feed_dict)


def measure_latency(cuda_session: CudaSession, input_dict):
    start = time.time()
    _ = cuda_session.infer(input_dict)
    end = time.time()
    return end - start


def run_torch(config: TestConfig):
    device_type = config.device.type
    is_cuda = device_type == "cuda"

    # Turn on TF32 for Ampere GPUs which could help when data type is float32.
    if is_cuda and torch.cuda.get_device_properties(0).major >= 8 and config.use_tf32:
        torch.backends.cuda.matmul.allow_tf32 = True
        torch.backends.cudnn.allow_tf32 = True

    enabled_auto_cast = is_cuda and config.dtype != torch.float32
    ort_inputs = config.random_inputs()

    with torch.inference_mode(), torch.autocast(device_type=device_type, dtype=config.dtype, enabled=enabled_auto_cast):
        sam2_model = load_sam2_model(config.sam2_dir, config.model_type, device=config.device)
        if config.component == "image_encoder":
            if is_cuda and config.torch_compile_mode != "none":
                sam2_model.image_encoder.forward = torch.compile(
                    sam2_model.image_encoder.forward,
                    mode=config.torch_compile_mode,  # "reduce-overhead" if you want to reduce latency of first run.
                    fullgraph=True,
                    dynamic=False,
                )

            image_shape = config.shape_dict()["image"]
            img = torch.randn(image_shape).to(device=config.device, dtype=config.dtype)
            sam2_encoder = SAM2ImageEncoder(sam2_model)

            if is_cuda and config.torch_compile_mode != "none":
                print(f"Running warm up. It will take a while since torch compile mode is {config.torch_compile_mode}.")

            for _ in range(config.warm_up):
                _image_features_0, _image_features_1, _image_embeddings = sam2_encoder(img)

            if is_cuda and config.enable_nvtx_profile:
                import nvtx
                from cuda import cudart

                cudart.cudaProfilerStart()
                print("Start nvtx profiling on encoder ...")
                with nvtx.annotate("one_run"):
                    sam2_encoder(img, enable_nvtx_profile=True)
                cudart.cudaProfilerStop()

            if is_cuda and config.enable_torch_profile:
                with torch.profiler.profile(
                    activities=[torch.profiler.ProfilerActivity.CPU, torch.profiler.ProfilerActivity.CUDA],
                    record_shapes=True,
                ) as prof:
                    print("Start torch profiling on encoder ...")
                    with torch.profiler.record_function("encoder"):
                        sam2_encoder(img)
                print(prof.key_averages().table(sort_by="cuda_time_total", row_limit=10))
                prof.export_chrome_trace("torch_image_encoder.json")

            if config.repeats == 0:
                return

            print(f"Start {config.repeats} runs of performance tests...")
            start = time.time()
            for _ in range(config.repeats):
                _image_features_0, _image_features_1, _image_embeddings = sam2_encoder(img)
                if is_cuda:
                    torch.cuda.synchronize()
        else:
            torch_inputs = (
                ort_inputs["image_features_0"],
                ort_inputs["image_features_1"],
                ort_inputs["image_embeddings"],
                ort_inputs["point_coords"],
                ort_inputs["point_labels"],
                ort_inputs["input_masks"],
                ort_inputs["has_input_masks"],
                ort_inputs["original_image_size"],
            )

            sam2_decoder = SAM2ImageDecoder(
                sam2_model,
                multimask_output=config.multi_mask_output,
            )

            if is_cuda and config.torch_compile_mode != "none":
                sam2_decoder.forward = torch.compile(
                    sam2_decoder.forward,
                    mode=config.torch_compile_mode,
                    fullgraph=True,
                    dynamic=False,
                )

            # warm up
            for _ in range(config.warm_up):
                _masks, _iou_predictions, _low_res_masks = sam2_decoder(*torch_inputs)

            if is_cuda and config.enable_nvtx_profile:
                import nvtx
                from cuda import cudart

                cudart.cudaProfilerStart()
                print("Start nvtx profiling on decoder...")
                with nvtx.annotate("one_run"):
                    sam2_decoder(*torch_inputs, enable_nvtx_profile=True)
                cudart.cudaProfilerStop()

            if is_cuda and config.enable_torch_profile:
                with torch.profiler.profile(
                    activities=[torch.profiler.ProfilerActivity.CPU, torch.profiler.ProfilerActivity.CUDA],
                    record_shapes=True,
                ) as prof:
                    print("Start torch profiling on decoder ...")
                    with torch.profiler.record_function("decoder"):
                        sam2_decoder(*torch_inputs)
                print(prof.key_averages().table(sort_by="cuda_time_total", row_limit=10))
                prof.export_chrome_trace("torch_image_decoder.json")

            if config.repeats == 0:
                return

            print(f"Start {config.repeats} runs of performance tests...")
            start = time.time()
            for _ in range(config.repeats):
                _masks, _iou_predictions, _low_res_masks = sam2_decoder(*torch_inputs)
                if is_cuda:
                    torch.cuda.synchronize()

        end = time.time()
        return (end - start) / config.repeats


def run_test(
    args: argparse.Namespace,
    csv_writer: csv.DictWriter | None = None,
):
    use_gpu: bool = args.use_gpu
    enable_cuda_graph: bool = args.use_cuda_graph
    repeats: int = args.repeats

    if use_gpu:
        device_id = torch.cuda.current_device()
        device = torch.device("cuda", device_id)
        provider = "CUDAExecutionProvider"
    else:
        device_id = 0
        device = torch.device("cpu")
        enable_cuda_graph = False
        provider = "CPUExecutionProvider"

    dtypes = {"fp32": torch.float32, "fp16": torch.float16, "bf16": torch.bfloat16}
    config = TestConfig(
        model_type=args.model_type,
        onnx_path=args.onnx_path,
        sam2_dir=args.sam2_dir,
        component=args.component,
        provider=provider,
        batch_size=args.batch_size,
        height=args.height,
        width=args.width,
        device=device,
        use_tf32=True,
        enable_cuda_graph=enable_cuda_graph,
        dtype=dtypes[args.dtype],
        prefer_nhwc=args.prefer_nhwc,
        repeats=args.repeats,
        warm_up=args.warm_up,
        enable_nvtx_profile=args.enable_nvtx_profile,
        enable_ort_profile=args.enable_ort_profile,
        enable_torch_profile=args.enable_torch_profile,
        torch_compile_mode=args.torch_compile_mode,
        verbose=False,
    )

    if args.engine == "ort":
        sess_options = SessionOptions()
        sess_options.intra_op_num_threads = args.intra_op_num_threads
        if config.enable_ort_profile:
            sess_options.enable_profiling = True
            sess_options.log_severity_level = 4
            sess_options.log_verbosity_level = 0

        session = create_session(config, sess_options)
        input_dict = config.random_inputs()

        # warm up session
        try:
            for _ in range(config.warm_up):
                _ = measure_latency(session, input_dict)
        except Exception as e:
            print(f"Failed to run {config=}. Exception: {e}")
            return

        if config.enable_nvtx_profile:
            import nvtx
            from cuda import cudart

            cudart.cudaProfilerStart()
            with nvtx.annotate("one_run"):
                _ = session.infer(input_dict)
            cudart.cudaProfilerStop()

        if config.enable_ort_profile:
            session.ort_session.end_profiling()

        if repeats == 0:
            return

        latency_list = []
        for _ in range(repeats):
            latency = measure_latency(session, input_dict)
            latency_list.append(latency)
        average_latency = statistics.mean(latency_list)

        del session
    else:  # torch
        with torch.no_grad():
            try:
                average_latency = run_torch(config)
            except Exception as e:
                print(f"Failed to run {config=}. Exception: {e}")
                return

        if repeats == 0:
            return

    engine = args.engine + ":" + ("cuda" if use_gpu else "cpu")
    row = {
        "model_type": args.model_type,
        "component": args.component,
        "dtype": args.dtype,
        "use_gpu": use_gpu,
        "enable_cuda_graph": enable_cuda_graph,
        "prefer_nhwc": config.prefer_nhwc,
        "use_tf32": config.use_tf32,
        "batch_size": args.batch_size,
        "height": args.height,
        "width": args.width,
        "multi_mask_output": args.multimask_output,
        "num_labels": config.num_labels,
        "num_points": config.num_points,
        "num_masks": config.num_masks,
        "intra_op_num_threads": args.intra_op_num_threads,
        "warm_up": config.warm_up,
        "repeats": repeats,
        "enable_nvtx_profile": args.enable_nvtx_profile,
        "torch_compile_mode": args.torch_compile_mode,
        "engine": engine,
        "average_latency": average_latency,
    }

    if csv_writer is not None:
        csv_writer.writerow(row)

    print(f"{vars(config)}")
    print(f"{row}")


def run_perf_test(args):
    features = "gpu" if args.use_gpu else "cpu"
    csv_filename = "benchmark_sam_{}_{}_{}.csv".format(
        features,
        args.engine,
        datetime.now().strftime("%Y%m%d-%H%M%S"),
    )
    with open(csv_filename, mode="a", newline="") as csv_file:
        column_names = [
            "model_type",
            "component",
            "dtype",
            "use_gpu",
            "enable_cuda_graph",
            "prefer_nhwc",
            "use_tf32",
            "batch_size",
            "height",
            "width",
            "multi_mask_output",
            "num_labels",
            "num_points",
            "num_masks",
            "intra_op_num_threads",
            "warm_up",
            "repeats",
            "enable_nvtx_profile",
            "torch_compile_mode",
            "engine",
            "average_latency",
        ]
        csv_writer = csv.DictWriter(csv_file, fieldnames=column_names)
        csv_writer.writeheader()

        run_test(args, csv_writer)


def _parse_arguments():
    parser = argparse.ArgumentParser(description="Benchmark SMA2 for ONNX Runtime and PyTorch.")

    parser.add_argument(
        "--component",
        required=False,
        choices=["image_encoder", "image_decoder"],
        default="image_encoder",
        help="component to benchmark. Choices are image_encoder and image_decoder.",
    )

    parser.add_argument(
        "--dtype", required=False, choices=["fp32", "fp16", "bf16"], default="fp32", help="Data type for inference."
    )

    parser.add_argument(
        "--use_gpu",
        required=False,
        action="store_true",
        help="Use GPU for inference.",
    )
    parser.set_defaults(use_gpu=False)

    parser.add_argument(
        "--use_cuda_graph",
        required=False,
        action="store_true",
        help="Use cuda graph in onnxruntime.",
    )
    parser.set_defaults(use_cuda_graph=False)

    parser.add_argument(
        "--intra_op_num_threads",
        required=False,
        type=int,
        choices=[0, 1, 2, 4, 8, 16],
        default=0,
        help="intra_op_num_threads for onnxruntime. ",
    )

    parser.add_argument(
        "--batch_size",
        required=False,
        type=int,
        default=1,
        help="batch size",
    )

    parser.add_argument(
        "--height",
        required=False,
        type=int,
        default=1024,
        help="image height",
    )

    parser.add_argument(
        "--width",
        required=False,
        type=int,
        default=1024,
        help="image width",
    )

    parser.add_argument(
        "--repeats",
        required=False,
        type=int,
        default=1000,
        help="number of repeats for performance test. Default is 1000.",
    )

    parser.add_argument(
        "--warm_up",
        required=False,
        type=int,
        default=5,
        help="number of runs for warm up. Default is 5.",
    )

    parser.add_argument(
        "--engine",
        required=False,
        type=str,
        default="ort",
        choices=["ort", "torch"],
        help="engine for inference",
    )

    parser.add_argument(
        "--multimask_output",
        required=False,
        default=False,
        action="store_true",
        help="Export mask_decoder or image_decoder with multimask_output",
    )

    parser.add_argument(
        "--prefer_nhwc",
        required=False,
        default=False,
        action="store_true",
        help="Use prefer_nhwc=1 provider option for CUDAExecutionProvider",
    )

    parser.add_argument(
        "--enable_nvtx_profile",
        required=False,
        default=False,
        action="store_true",
        help="Enable nvtx profiling. It will add an extra run for profiling before performance test.",
    )

    parser.add_argument(
        "--enable_ort_profile",
        required=False,
        default=False,
        action="store_true",
        help="Enable ORT profiling.",
    )

    parser.add_argument(
        "--enable_torch_profile",
        required=False,
        default=False,
        action="store_true",
        help="Enable PyTorch profiling. It will add an extra run for profiling before performance test.",
    )

    parser.add_argument(
        "--model_type",
        required=False,
        type=str,
        default="sam2_hiera_large",
        choices=["sam2_hiera_tiny", "sam2_hiera_small", "sam2_hiera_large", "sam2_hiera_base_plus"],
        help="sam2 model name",
    )

    parser.add_argument(
        "--sam2_dir",
        required=False,
        type=str,
        default="./segment-anything-2",
        help="The directory of segment-anything-2 git root directory",
    )

    parser.add_argument(
        "--onnx_path",
        required=False,
        type=str,
        default="./sam2_onnx_models/sam2_hiera_large_image_encoder.onnx",
        help="path of onnx model",
    )

    parser.add_argument(
        "--torch_compile_mode",
        required=False,
        type=str,
        default=None,
        choices=["reduce-overhead", "max-autotune", "max-autotune-no-cudagraphs", "none"],
        help="torch compile mode. none will disable torch compile.",
    )

    args = parser.parse_args()

    return args


if __name__ == "__main__":
    args = _parse_arguments()
    print(f"arguments:{args}")

    if args.torch_compile_mode is None:
        # image decoder will fail with compile modes other than "none".
        args.torch_compile_mode = "max-autotune" if args.component == "image_encoder" else "none"

    if args.use_gpu:
        assert torch.cuda.is_available()
        if args.engine == "ort":
            assert "CUDAExecutionProvider" in get_available_providers()
            args.enable_torch_profile = False
    else:
        # Only support cuda profiling for now.
        assert not args.enable_nvtx_profile
        assert not args.enable_torch_profile

    if args.enable_nvtx_profile or args.enable_torch_profile:
        run_test(args)
    else:
        run_perf_test(args)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\reshape\tile.py ===
"""
Quantilization functions and related stuff
"""
from __future__ import annotations

from typing import (
    TYPE_CHECKING,
    Any,
    Callable,
    Literal,
)

import numpy as np

from pandas._libs import (
    Timedelta,
    Timestamp,
    lib,
)

from pandas.core.dtypes.common import (
    ensure_platform_int,
    is_bool_dtype,
    is_integer,
    is_list_like,
    is_numeric_dtype,
    is_scalar,
)
from pandas.core.dtypes.dtypes import (
    CategoricalDtype,
    DatetimeTZDtype,
    ExtensionDtype,
)
from pandas.core.dtypes.generic import ABCSeries
from pandas.core.dtypes.missing import isna

from pandas import (
    Categorical,
    Index,
    IntervalIndex,
)
import pandas.core.algorithms as algos
from pandas.core.arrays.datetimelike import dtype_to_unit

if TYPE_CHECKING:
    from pandas._typing import (
        DtypeObj,
        IntervalLeftRight,
    )


def cut(
    x,
    bins,
    right: bool = True,
    labels=None,
    retbins: bool = False,
    precision: int = 3,
    include_lowest: bool = False,
    duplicates: str = "raise",
    ordered: bool = True,
):
    """
    Bin values into discrete intervals.

    Use `cut` when you need to segment and sort data values into bins. This
    function is also useful for going from a continuous variable to a
    categorical variable. For example, `cut` could convert ages to groups of
    age ranges. Supports binning into an equal number of bins, or a
    pre-specified array of bins.

    Parameters
    ----------
    x : array-like
        The input array to be binned. Must be 1-dimensional.
    bins : int, sequence of scalars, or IntervalIndex
        The criteria to bin by.

        * int : Defines the number of equal-width bins in the range of `x`. The
          range of `x` is extended by .1% on each side to include the minimum
          and maximum values of `x`.
        * sequence of scalars : Defines the bin edges allowing for non-uniform
          width. No extension of the range of `x` is done.
        * IntervalIndex : Defines the exact bins to be used. Note that
          IntervalIndex for `bins` must be non-overlapping.

    right : bool, default True
        Indicates whether `bins` includes the rightmost edge or not. If
        ``right == True`` (the default), then the `bins` ``[1, 2, 3, 4]``
        indicate (1,2], (2,3], (3,4]. This argument is ignored when
        `bins` is an IntervalIndex.
    labels : array or False, default None
        Specifies the labels for the returned bins. Must be the same length as
        the resulting bins. If False, returns only integer indicators of the
        bins. This affects the type of the output container (see below).
        This argument is ignored when `bins` is an IntervalIndex. If True,
        raises an error. When `ordered=False`, labels must be provided.
    retbins : bool, default False
        Whether to return the bins or not. Useful when bins is provided
        as a scalar.
    precision : int, default 3
        The precision at which to store and display the bins labels.
    include_lowest : bool, default False
        Whether the first interval should be left-inclusive or not.
    duplicates : {default 'raise', 'drop'}, optional
        If bin edges are not unique, raise ValueError or drop non-uniques.
    ordered : bool, default True
        Whether the labels are ordered or not. Applies to returned types
        Categorical and Series (with Categorical dtype). If True,
        the resulting categorical will be ordered. If False, the resulting
        categorical will be unordered (labels must be provided).

    Returns
    -------
    out : Categorical, Series, or ndarray
        An array-like object representing the respective bin for each value
        of `x`. The type depends on the value of `labels`.

        * None (default) : returns a Series for Series `x` or a
          Categorical for all other inputs. The values stored within
          are Interval dtype.

        * sequence of scalars : returns a Series for Series `x` or a
          Categorical for all other inputs. The values stored within
          are whatever the type in the sequence is.

        * False : returns an ndarray of integers.

    bins : numpy.ndarray or IntervalIndex.
        The computed or specified bins. Only returned when `retbins=True`.
        For scalar or sequence `bins`, this is an ndarray with the computed
        bins. If set `duplicates=drop`, `bins` will drop non-unique bin. For
        an IntervalIndex `bins`, this is equal to `bins`.

    See Also
    --------
    qcut : Discretize variable into equal-sized buckets based on rank
        or based on sample quantiles.
    Categorical : Array type for storing data that come from a
        fixed set of values.
    Series : One-dimensional array with axis labels (including time series).
    IntervalIndex : Immutable Index implementing an ordered, sliceable set.

    Notes
    -----
    Any NA values will be NA in the result. Out of bounds values will be NA in
    the resulting Series or Categorical object.

    Reference :ref:`the user guide <reshaping.tile.cut>` for more examples.

    Examples
    --------
    Discretize into three equal-sized bins.

    >>> pd.cut(np.array([1, 7, 5, 4, 6, 3]), 3)
    ... # doctest: +ELLIPSIS
    [(0.994, 3.0], (5.0, 7.0], (3.0, 5.0], (3.0, 5.0], (5.0, 7.0], ...
    Categories (3, interval[float64, right]): [(0.994, 3.0] < (3.0, 5.0] ...

    >>> pd.cut(np.array([1, 7, 5, 4, 6, 3]), 3, retbins=True)
    ... # doctest: +ELLIPSIS
    ([(0.994, 3.0], (5.0, 7.0], (3.0, 5.0], (3.0, 5.0], (5.0, 7.0], ...
    Categories (3, interval[float64, right]): [(0.994, 3.0] < (3.0, 5.0] ...
    array([0.994, 3.   , 5.   , 7.   ]))

    Discovers the same bins, but assign them specific labels. Notice that
    the returned Categorical's categories are `labels` and is ordered.

    >>> pd.cut(np.array([1, 7, 5, 4, 6, 3]),
    ...        3, labels=["bad", "medium", "good"])
    ['bad', 'good', 'medium', 'medium', 'good', 'bad']
    Categories (3, object): ['bad' < 'medium' < 'good']

    ``ordered=False`` will result in unordered categories when labels are passed.
    This parameter can be used to allow non-unique labels:

    >>> pd.cut(np.array([1, 7, 5, 4, 6, 3]), 3,
    ...        labels=["B", "A", "B"], ordered=False)
    ['B', 'B', 'A', 'A', 'B', 'B']
    Categories (2, object): ['A', 'B']

    ``labels=False`` implies you just want the bins back.

    >>> pd.cut([0, 1, 1, 2], bins=4, labels=False)
    array([0, 1, 1, 3])

    Passing a Series as an input returns a Series with categorical dtype:

    >>> s = pd.Series(np.array([2, 4, 6, 8, 10]),
    ...               index=['a', 'b', 'c', 'd', 'e'])
    >>> pd.cut(s, 3)
    ... # doctest: +ELLIPSIS
    a    (1.992, 4.667]
    b    (1.992, 4.667]
    c    (4.667, 7.333]
    d     (7.333, 10.0]
    e     (7.333, 10.0]
    dtype: category
    Categories (3, interval[float64, right]): [(1.992, 4.667] < (4.667, ...

    Passing a Series as an input returns a Series with mapping value.
    It is used to map numerically to intervals based on bins.

    >>> s = pd.Series(np.array([2, 4, 6, 8, 10]),
    ...               index=['a', 'b', 'c', 'd', 'e'])
    >>> pd.cut(s, [0, 2, 4, 6, 8, 10], labels=False, retbins=True, right=False)
    ... # doctest: +ELLIPSIS
    (a    1.0
     b    2.0
     c    3.0
     d    4.0
     e    NaN
     dtype: float64,
     array([ 0,  2,  4,  6,  8, 10]))

    Use `drop` optional when bins is not unique

    >>> pd.cut(s, [0, 2, 4, 6, 10, 10], labels=False, retbins=True,
    ...        right=False, duplicates='drop')
    ... # doctest: +ELLIPSIS
    (a    1.0
     b    2.0
     c    3.0
     d    3.0
     e    NaN
     dtype: float64,
     array([ 0,  2,  4,  6, 10]))

    Passing an IntervalIndex for `bins` results in those categories exactly.
    Notice that values not covered by the IntervalIndex are set to NaN. 0
    is to the left of the first bin (which is closed on the right), and 1.5
    falls between two bins.

    >>> bins = pd.IntervalIndex.from_tuples([(0, 1), (2, 3), (4, 5)])
    >>> pd.cut([0, 0.5, 1.5, 2.5, 4.5], bins)
    [NaN, (0.0, 1.0], NaN, (2.0, 3.0], (4.0, 5.0]]
    Categories (3, interval[int64, right]): [(0, 1] < (2, 3] < (4, 5]]
    """
    # NOTE: this binning code is changed a bit from histogram for var(x) == 0

    original = x
    x_idx = _preprocess_for_cut(x)
    x_idx, _ = _coerce_to_type(x_idx)

    if not np.iterable(bins):
        bins = _nbins_to_bins(x_idx, bins, right)

    elif isinstance(bins, IntervalIndex):
        if bins.is_overlapping:
            raise ValueError("Overlapping IntervalIndex is not accepted.")

    else:
        bins = Index(bins)
        if not bins.is_monotonic_increasing:
            raise ValueError("bins must increase monotonically.")

    fac, bins = _bins_to_cuts(
        x_idx,
        bins,
        right=right,
        labels=labels,
        precision=precision,
        include_lowest=include_lowest,
        duplicates=duplicates,
        ordered=ordered,
    )

    return _postprocess_for_cut(fac, bins, retbins, original)


def qcut(
    x,
    q,
    labels=None,
    retbins: bool = False,
    precision: int = 3,
    duplicates: str = "raise",
):
    """
    Quantile-based discretization function.

    Discretize variable into equal-sized buckets based on rank or based
    on sample quantiles. For example 1000 values for 10 quantiles would
    produce a Categorical object indicating quantile membership for each data point.

    Parameters
    ----------
    x : 1d ndarray or Series
    q : int or list-like of float
        Number of quantiles. 10 for deciles, 4 for quartiles, etc. Alternately
        array of quantiles, e.g. [0, .25, .5, .75, 1.] for quartiles.
    labels : array or False, default None
        Used as labels for the resulting bins. Must be of the same length as
        the resulting bins. If False, return only integer indicators of the
        bins. If True, raises an error.
    retbins : bool, optional
        Whether to return the (bins, labels) or not. Can be useful if bins
        is given as a scalar.
    precision : int, optional
        The precision at which to store and display the bins labels.
    duplicates : {default 'raise', 'drop'}, optional
        If bin edges are not unique, raise ValueError or drop non-uniques.

    Returns
    -------
    out : Categorical or Series or array of integers if labels is False
        The return type (Categorical or Series) depends on the input: a Series
        of type category if input is a Series else Categorical. Bins are
        represented as categories when categorical data is returned.
    bins : ndarray of floats
        Returned only if `retbins` is True.

    Notes
    -----
    Out of bounds values will be NA in the resulting Categorical object

    Examples
    --------
    >>> pd.qcut(range(5), 4)
    ... # doctest: +ELLIPSIS
    [(-0.001, 1.0], (-0.001, 1.0], (1.0, 2.0], (2.0, 3.0], (3.0, 4.0]]
    Categories (4, interval[float64, right]): [(-0.001, 1.0] < (1.0, 2.0] ...

    >>> pd.qcut(range(5), 3, labels=["good", "medium", "bad"])
    ... # doctest: +SKIP
    [good, good, medium, bad, bad]
    Categories (3, object): [good < medium < bad]

    >>> pd.qcut(range(5), 4, labels=False)
    array([0, 0, 1, 2, 3])
    """
    original = x
    x_idx = _preprocess_for_cut(x)
    x_idx, _ = _coerce_to_type(x_idx)

    quantiles = np.linspace(0, 1, q + 1) if is_integer(q) else q

    bins = x_idx.to_series().dropna().quantile(quantiles)

    fac, bins = _bins_to_cuts(
        x_idx,
        Index(bins),
        labels=labels,
        precision=precision,
        include_lowest=True,
        duplicates=duplicates,
    )

    return _postprocess_for_cut(fac, bins, retbins, original)


def _nbins_to_bins(x_idx: Index, nbins: int, right: bool) -> Index:
    """
    If a user passed an integer N for bins, convert this to a sequence of N
    equal(ish)-sized bins.
    """
    if is_scalar(nbins) and nbins < 1:
        raise ValueError("`bins` should be a positive integer.")

    if x_idx.size == 0:
        raise ValueError("Cannot cut empty array")

    rng = (x_idx.min(), x_idx.max())
    mn, mx = rng

    if is_numeric_dtype(x_idx.dtype) and (np.isinf(mn) or np.isinf(mx)):
        # GH#24314
        raise ValueError(
            "cannot specify integer `bins` when input data contains infinity"
        )

    if mn == mx:  # adjust end points before binning
        if _is_dt_or_td(x_idx.dtype):
            # using seconds=1 is pretty arbitrary here
            # error: Argument 1 to "dtype_to_unit" has incompatible type
            # "dtype[Any] | ExtensionDtype"; expected "DatetimeTZDtype | dtype[Any]"
            unit = dtype_to_unit(x_idx.dtype)  # type: ignore[arg-type]
            td = Timedelta(seconds=1).as_unit(unit)
            # Use DatetimeArray/TimedeltaArray method instead of linspace
            # error: Item "ExtensionArray" of "ExtensionArray | ndarray[Any, Any]"
            # has no attribute "_generate_range"
            bins = x_idx._values._generate_range(  # type: ignore[union-attr]
                start=mn - td, end=mx + td, periods=nbins + 1, freq=None, unit=unit
            )
        else:
            mn -= 0.001 * abs(mn) if mn != 0 else 0.001
            mx += 0.001 * abs(mx) if mx != 0 else 0.001

            bins = np.linspace(mn, mx, nbins + 1, endpoint=True)
    else:  # adjust end points after binning
        if _is_dt_or_td(x_idx.dtype):
            # Use DatetimeArray/TimedeltaArray method instead of linspace

            # error: Argument 1 to "dtype_to_unit" has incompatible type
            # "dtype[Any] | ExtensionDtype"; expected "DatetimeTZDtype | dtype[Any]"
            unit = dtype_to_unit(x_idx.dtype)  # type: ignore[arg-type]
            # error: Item "ExtensionArray" of "ExtensionArray | ndarray[Any, Any]"
            # has no attribute "_generate_range"
            bins = x_idx._values._generate_range(  # type: ignore[union-attr]
                start=mn, end=mx, periods=nbins + 1, freq=None, unit=unit
            )
        else:
            bins = np.linspace(mn, mx, nbins + 1, endpoint=True)
        adj = (mx - mn) * 0.001  # 0.1% of the range
        if right:
            bins[0] -= adj
        else:
            bins[-1] += adj

    return Index(bins)


def _bins_to_cuts(
    x_idx: Index,
    bins: Index,
    right: bool = True,
    labels=None,
    precision: int = 3,
    include_lowest: bool = False,
    duplicates: str = "raise",
    ordered: bool = True,
):
    if not ordered and labels is None:
        raise ValueError("'labels' must be provided if 'ordered = False'")

    if duplicates not in ["raise", "drop"]:
        raise ValueError(
            "invalid value for 'duplicates' parameter, valid options are: raise, drop"
        )

    result: Categorical | np.ndarray

    if isinstance(bins, IntervalIndex):
        # we have a fast-path here
        ids = bins.get_indexer(x_idx)
        cat_dtype = CategoricalDtype(bins, ordered=True)
        result = Categorical.from_codes(ids, dtype=cat_dtype, validate=False)
        return result, bins

    unique_bins = algos.unique(bins)
    if len(unique_bins) < len(bins) and len(bins) != 2:
        if duplicates == "raise":
            raise ValueError(
                f"Bin edges must be unique: {repr(bins)}.\n"
                f"You can drop duplicate edges by setting the 'duplicates' kwarg"
            )
        bins = unique_bins

    side: Literal["left", "right"] = "left" if right else "right"

    try:
        ids = bins.searchsorted(x_idx, side=side)
    except TypeError as err:
        # e.g. test_datetime_nan_error if bins are DatetimeArray and x_idx
        #  is integers
        if x_idx.dtype.kind == "m":
            raise ValueError("bins must be of timedelta64 dtype") from err
        elif x_idx.dtype.kind == bins.dtype.kind == "M":
            raise ValueError(
                "Cannot use timezone-naive bins with timezone-aware values, "
                "or vice-versa"
            ) from err
        elif x_idx.dtype.kind == "M":
            raise ValueError("bins must be of datetime64 dtype") from err
        else:
            raise
    ids = ensure_platform_int(ids)

    if include_lowest:
        ids[x_idx == bins[0]] = 1

    na_mask = isna(x_idx) | (ids == len(bins)) | (ids == 0)
    has_nas = na_mask.any()

    if labels is not False:
        if not (labels is None or is_list_like(labels)):
            raise ValueError(
                "Bin labels must either be False, None or passed in as a "
                "list-like argument"
            )

        if labels is None:
            labels = _format_labels(
                bins, precision, right=right, include_lowest=include_lowest
            )
        elif ordered and len(set(labels)) != len(labels):
            raise ValueError(
                "labels must be unique if ordered=True; pass ordered=False "
                "for duplicate labels"
            )
        else:
            if len(labels) != len(bins) - 1:
                raise ValueError(
                    "Bin labels must be one fewer than the number of bin edges"
                )

        if not isinstance(getattr(labels, "dtype", None), CategoricalDtype):
            labels = Categorical(
                labels,
                categories=labels if len(set(labels)) == len(labels) else None,
                ordered=ordered,
            )
        # TODO: handle mismatch between categorical label order and pandas.cut order.
        np.putmask(ids, na_mask, 0)
        result = algos.take_nd(labels, ids - 1)

    else:
        result = ids - 1
        if has_nas:
            result = result.astype(np.float64)
            np.putmask(result, na_mask, np.nan)

    return result, bins


def _coerce_to_type(x: Index) -> tuple[Index, DtypeObj | None]:
    """
    if the passed data is of datetime/timedelta, bool or nullable int type,
    this method converts it to numeric so that cut or qcut method can
    handle it
    """
    dtype: DtypeObj | None = None

    if _is_dt_or_td(x.dtype):
        dtype = x.dtype
    elif is_bool_dtype(x.dtype):
        # GH 20303
        x = x.astype(np.int64)
    # To support cut and qcut for IntegerArray we convert to float dtype.
    # Will properly support in the future.
    # https://github.com/pandas-dev/pandas/pull/31290
    # https://github.com/pandas-dev/pandas/issues/31389
    elif isinstance(x.dtype, ExtensionDtype) and is_numeric_dtype(x.dtype):
        x_arr = x.to_numpy(dtype=np.float64, na_value=np.nan)
        x = Index(x_arr)

    return Index(x), dtype


def _is_dt_or_td(dtype: DtypeObj) -> bool:
    # Note: the dtype here comes from an Index.dtype, so we know that that any
    #  dt64/td64 dtype is of a supported unit.
    return isinstance(dtype, DatetimeTZDtype) or lib.is_np_dtype(dtype, "mM")


def _format_labels(
    bins: Index,
    precision: int,
    right: bool = True,
    include_lowest: bool = False,
):
    """based on the dtype, return our labels"""
    closed: IntervalLeftRight = "right" if right else "left"

    formatter: Callable[[Any], Timestamp] | Callable[[Any], Timedelta]

    if _is_dt_or_td(bins.dtype):
        # error: Argument 1 to "dtype_to_unit" has incompatible type
        # "dtype[Any] | ExtensionDtype"; expected "DatetimeTZDtype | dtype[Any]"
        unit = dtype_to_unit(bins.dtype)  # type: ignore[arg-type]
        formatter = lambda x: x
        adjust = lambda x: x - Timedelta(1, unit=unit).as_unit(unit)
    else:
        precision = _infer_precision(precision, bins)
        formatter = lambda x: _round_frac(x, precision)
        adjust = lambda x: x - 10 ** (-precision)

    breaks = [formatter(b) for b in bins]
    if right and include_lowest:
        # adjust lhs of first interval by precision to account for being right closed
        breaks[0] = adjust(breaks[0])

    if _is_dt_or_td(bins.dtype):
        # error: "Index" has no attribute "as_unit"
        breaks = type(bins)(breaks).as_unit(unit)  # type: ignore[attr-defined]

    return IntervalIndex.from_breaks(breaks, closed=closed)


def _preprocess_for_cut(x) -> Index:
    """
    handles preprocessing for cut where we convert passed
    input to array, strip the index information and store it
    separately
    """
    # Check that the passed array is a Pandas or Numpy object
    # We don't want to strip away a Pandas data-type here (e.g. datetimetz)
    ndim = getattr(x, "ndim", None)
    if ndim is None:
        x = np.asarray(x)
    if x.ndim != 1:
        raise ValueError("Input array must be 1 dimensional")

    return Index(x)


def _postprocess_for_cut(fac, bins, retbins: bool, original):
    """
    handles post processing for the cut method where
    we combine the index information if the originally passed
    datatype was a series
    """
    if isinstance(original, ABCSeries):
        fac = original._constructor(fac, index=original.index, name=original.name)

    if not retbins:
        return fac

    if isinstance(bins, Index) and is_numeric_dtype(bins.dtype):
        bins = bins._values

    return fac, bins


def _round_frac(x, precision: int):
    """
    Round the fractional part of the given number
    """
    if not np.isfinite(x) or x == 0:
        return x
    else:
        frac, whole = np.modf(x)
        if whole == 0:
            digits = -int(np.floor(np.log10(abs(frac)))) - 1 + precision
        else:
            digits = precision
        return np.around(x, digits)


def _infer_precision(base_precision: int, bins: Index) -> int:
    """
    Infer an appropriate precision for _round_frac
    """
    for precision in range(base_precision, 20):
        levels = np.asarray([_round_frac(b, precision) for b in bins])
        if algos.unique(levels).size == bins.size:
            return precision
    return base_precision  # default

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\domains\algebraicfield.py ===
"""Implementation of :class:`AlgebraicField` class. """


from sympy.core.add import Add
from sympy.core.mul import Mul
from sympy.core.singleton import S
from sympy.core.symbol import Dummy, symbols
from sympy.polys.domains.characteristiczero import CharacteristicZero
from sympy.polys.domains.field import Field
from sympy.polys.domains.simpledomain import SimpleDomain
from sympy.polys.polyclasses import ANP
from sympy.polys.polyerrors import CoercionFailed, DomainError, NotAlgebraic, IsomorphismFailed
from sympy.utilities import public

@public
class AlgebraicField(Field, CharacteristicZero, SimpleDomain):
    r"""Algebraic number field :ref:`QQ(a)`

    A :ref:`QQ(a)` domain represents an `algebraic number field`_
    `\mathbb{Q}(a)` as a :py:class:`~.Domain` in the domain system (see
    :ref:`polys-domainsintro`).

    A :py:class:`~.Poly` created from an expression involving `algebraic
    numbers`_ will treat the algebraic numbers as generators if the generators
    argument is not specified.

    >>> from sympy import Poly, Symbol, sqrt
    >>> x = Symbol('x')
    >>> Poly(x**2 + sqrt(2))
    Poly(x**2 + (sqrt(2)), x, sqrt(2), domain='ZZ')

    That is a multivariate polynomial with ``sqrt(2)`` treated as one of the
    generators (variables). If the generators are explicitly specified then
    ``sqrt(2)`` will be considered to be a coefficient but by default the
    :ref:`EX` domain is used. To make a :py:class:`~.Poly` with a :ref:`QQ(a)`
    domain the argument ``extension=True`` can be given.

    >>> Poly(x**2 + sqrt(2), x)
    Poly(x**2 + sqrt(2), x, domain='EX')
    >>> Poly(x**2 + sqrt(2), x, extension=True)
    Poly(x**2 + sqrt(2), x, domain='QQ<sqrt(2)>')

    A generator of the algebraic field extension can also be specified
    explicitly which is particularly useful if the coefficients are all
    rational but an extension field is needed (e.g. to factor the
    polynomial).

    >>> Poly(x**2 + 1)
    Poly(x**2 + 1, x, domain='ZZ')
    >>> Poly(x**2 + 1, extension=sqrt(2))
    Poly(x**2 + 1, x, domain='QQ<sqrt(2)>')

    It is possible to factorise a polynomial over a :ref:`QQ(a)` domain using
    the ``extension`` argument to :py:func:`~.factor` or by specifying the domain
    explicitly.

    >>> from sympy import factor, QQ
    >>> factor(x**2 - 2)
    x**2 - 2
    >>> factor(x**2 - 2, extension=sqrt(2))
    (x - sqrt(2))*(x + sqrt(2))
    >>> factor(x**2 - 2, domain='QQ<sqrt(2)>')
    (x - sqrt(2))*(x + sqrt(2))
    >>> factor(x**2 - 2, domain=QQ.algebraic_field(sqrt(2)))
    (x - sqrt(2))*(x + sqrt(2))

    The ``extension=True`` argument can be used but will only create an
    extension that contains the coefficients which is usually not enough to
    factorise the polynomial.

    >>> p = x**3 + sqrt(2)*x**2 - 2*x - 2*sqrt(2)
    >>> factor(p)                         # treats sqrt(2) as a symbol
    (x + sqrt(2))*(x**2 - 2)
    >>> factor(p, extension=True)
    (x - sqrt(2))*(x + sqrt(2))**2
    >>> factor(x**2 - 2, extension=True)  # all rational coefficients
    x**2 - 2

    It is also possible to use :ref:`QQ(a)` with the :py:func:`~.cancel`
    and :py:func:`~.gcd` functions.

    >>> from sympy import cancel, gcd
    >>> cancel((x**2 - 2)/(x - sqrt(2)))
    (x**2 - 2)/(x - sqrt(2))
    >>> cancel((x**2 - 2)/(x - sqrt(2)), extension=sqrt(2))
    x + sqrt(2)
    >>> gcd(x**2 - 2, x - sqrt(2))
    1
    >>> gcd(x**2 - 2, x - sqrt(2), extension=sqrt(2))
    x - sqrt(2)

    When using the domain directly :ref:`QQ(a)` can be used as a constructor
    to create instances which then support the operations ``+,-,*,**,/``. The
    :py:meth:`~.Domain.algebraic_field` method is used to construct a
    particular :ref:`QQ(a)` domain. The :py:meth:`~.Domain.from_sympy` method
    can be used to create domain elements from normal SymPy expressions.

    >>> K = QQ.algebraic_field(sqrt(2))
    >>> K
    QQ<sqrt(2)>
    >>> xk = K.from_sympy(3 + 4*sqrt(2))
    >>> xk  # doctest: +SKIP
    ANP([4, 3], [1, 0, -2], QQ)

    Elements of :ref:`QQ(a)` are instances of :py:class:`~.ANP` which have
    limited printing support. The raw display shows the internal
    representation of the element as the list ``[4, 3]`` representing the
    coefficients of ``1`` and ``sqrt(2)`` for this element in the form
    ``a * sqrt(2) + b * 1`` where ``a`` and ``b`` are elements of :ref:`QQ`.
    The minimal polynomial for the generator ``(x**2 - 2)`` is also shown in
    the :ref:`dup-representation` as the list ``[1, 0, -2]``. We can use
    :py:meth:`~.Domain.to_sympy` to get a better printed form for the
    elements and to see the results of operations.

    >>> xk = K.from_sympy(3 + 4*sqrt(2))
    >>> yk = K.from_sympy(2 + 3*sqrt(2))
    >>> xk * yk  # doctest: +SKIP
    ANP([17, 30], [1, 0, -2], QQ)
    >>> K.to_sympy(xk * yk)
    17*sqrt(2) + 30
    >>> K.to_sympy(xk + yk)
    5 + 7*sqrt(2)
    >>> K.to_sympy(xk ** 2)
    24*sqrt(2) + 41
    >>> K.to_sympy(xk / yk)
    sqrt(2)/14 + 9/7

    Any expression representing an algebraic number can be used to generate
    a :ref:`QQ(a)` domain provided its `minimal polynomial`_ can be computed.
    The function :py:func:`~.minpoly` function is used for this.

    >>> from sympy import exp, I, pi, minpoly
    >>> g = exp(2*I*pi/3)
    >>> g
    exp(2*I*pi/3)
    >>> g.is_algebraic
    True
    >>> minpoly(g, x)
    x**2 + x + 1
    >>> factor(x**3 - 1, extension=g)
    (x - 1)*(x - exp(2*I*pi/3))*(x + 1 + exp(2*I*pi/3))

    It is also possible to make an algebraic field from multiple extension
    elements.

    >>> K = QQ.algebraic_field(sqrt(2), sqrt(3))
    >>> K
    QQ<sqrt(2) + sqrt(3)>
    >>> p = x**4 - 5*x**2 + 6
    >>> factor(p)
    (x**2 - 3)*(x**2 - 2)
    >>> factor(p, domain=K)
    (x - sqrt(2))*(x + sqrt(2))*(x - sqrt(3))*(x + sqrt(3))
    >>> factor(p, extension=[sqrt(2), sqrt(3)])
    (x - sqrt(2))*(x + sqrt(2))*(x - sqrt(3))*(x + sqrt(3))

    Multiple extension elements are always combined together to make a single
    `primitive element`_. In the case of ``[sqrt(2), sqrt(3)]`` the primitive
    element chosen is ``sqrt(2) + sqrt(3)`` which is why the domain displays
    as ``QQ<sqrt(2) + sqrt(3)>``. The minimal polynomial for the primitive
    element is computed using the :py:func:`~.primitive_element` function.

    >>> from sympy import primitive_element
    >>> primitive_element([sqrt(2), sqrt(3)], x)
    (x**4 - 10*x**2 + 1, [1, 1])
    >>> minpoly(sqrt(2) + sqrt(3), x)
    x**4 - 10*x**2 + 1

    The extension elements that generate the domain can be accessed from the
    domain using the :py:attr:`~.ext` and :py:attr:`~.orig_ext` attributes as
    instances of :py:class:`~.AlgebraicNumber`. The minimal polynomial for
    the primitive element as a :py:class:`~.DMP` instance is available as
    :py:attr:`~.mod`.

    >>> K = QQ.algebraic_field(sqrt(2), sqrt(3))
    >>> K
    QQ<sqrt(2) + sqrt(3)>
    >>> K.ext
    sqrt(2) + sqrt(3)
    >>> K.orig_ext
    (sqrt(2), sqrt(3))
    >>> K.mod  # doctest: +SKIP
    DMP_Python([1, 0, -10, 0, 1], QQ)

    The `discriminant`_ of the field can be obtained from the
    :py:meth:`~.discriminant` method, and an `integral basis`_ from the
    :py:meth:`~.integral_basis` method. The latter returns a list of
    :py:class:`~.ANP` instances by default, but can be made to return instances
    of :py:class:`~.Expr` or :py:class:`~.AlgebraicNumber` by passing a ``fmt``
    argument. The maximal order, or ring of integers, of the field can also be
    obtained from the :py:meth:`~.maximal_order` method, as a
    :py:class:`~sympy.polys.numberfields.modules.Submodule`.

    >>> zeta5 = exp(2*I*pi/5)
    >>> K = QQ.algebraic_field(zeta5)
    >>> K
    QQ<exp(2*I*pi/5)>
    >>> K.discriminant()
    125
    >>> K = QQ.algebraic_field(sqrt(5))
    >>> K
    QQ<sqrt(5)>
    >>> K.integral_basis(fmt='sympy')
    [1, 1/2 + sqrt(5)/2]
    >>> K.maximal_order()
    Submodule[[2, 0], [1, 1]]/2

    The factorization of a rational prime into prime ideals of the field is
    computed by the :py:meth:`~.primes_above` method, which returns a list
    of :py:class:`~sympy.polys.numberfields.primes.PrimeIdeal` instances.

    >>> zeta7 = exp(2*I*pi/7)
    >>> K = QQ.algebraic_field(zeta7)
    >>> K
    QQ<exp(2*I*pi/7)>
    >>> K.primes_above(11)
    [(11, _x**3 + 5*_x**2 + 4*_x - 1), (11, _x**3 - 4*_x**2 - 5*_x - 1)]

    The Galois group of the Galois closure of the field can be computed (when
    the minimal polynomial of the field is of sufficiently small degree).

    >>> K.galois_group(by_name=True)[0]
    S6TransitiveSubgroups.C6

    Notes
    =====

    It is not currently possible to generate an algebraic extension over any
    domain other than :ref:`QQ`. Ideally it would be possible to generate
    extensions like ``QQ(x)(sqrt(x**2 - 2))``. This is equivalent to the
    quotient ring ``QQ(x)[y]/(y**2 - x**2 + 2)`` and there are two
    implementations of this kind of quotient ring/extension in the
    :py:class:`~.QuotientRing` and :py:class:`~.MonogenicFiniteExtension`
    classes.  Each of those implementations needs some work to make them fully
    usable though.

    .. _algebraic number field: https://en.wikipedia.org/wiki/Algebraic_number_field
    .. _algebraic numbers: https://en.wikipedia.org/wiki/Algebraic_number
    .. _discriminant: https://en.wikipedia.org/wiki/Discriminant_of_an_algebraic_number_field
    .. _integral basis: https://en.wikipedia.org/wiki/Algebraic_number_field#Integral_basis
    .. _minimal polynomial: https://en.wikipedia.org/wiki/Minimal_polynomial_(field_theory)
    .. _primitive element: https://en.wikipedia.org/wiki/Primitive_element_theorem
    """

    dtype = ANP

    is_AlgebraicField = is_Algebraic = True
    is_Numerical = True

    has_assoc_Ring = False
    has_assoc_Field = True

    def __init__(self, dom, *ext, alias=None):
        r"""
        Parameters
        ==========

        dom : :py:class:`~.Domain`
            The base field over which this is an extension field.
            Currently only :ref:`QQ` is accepted.

        *ext : One or more :py:class:`~.Expr`
            Generators of the extension. These should be expressions that are
            algebraic over `\mathbb{Q}`.

        alias : str, :py:class:`~.Symbol`, None, optional (default=None)
            If provided, this will be used as the alias symbol for the
            primitive element of the :py:class:`~.AlgebraicField`.
            If ``None``, while ``ext`` consists of exactly one
            :py:class:`~.AlgebraicNumber`, its alias (if any) will be used.
        """
        if not dom.is_QQ:
            raise DomainError("ground domain must be a rational field")

        from sympy.polys.numberfields import to_number_field
        if len(ext) == 1 and isinstance(ext[0], tuple):
            orig_ext = ext[0][1:]
        else:
            orig_ext = ext

        if alias is None and len(ext) == 1:
            alias = getattr(ext[0], 'alias', None)

        self.orig_ext = orig_ext
        """
        Original elements given to generate the extension.

        >>> from sympy import QQ, sqrt
        >>> K = QQ.algebraic_field(sqrt(2), sqrt(3))
        >>> K.orig_ext
        (sqrt(2), sqrt(3))
        """

        self.ext = to_number_field(ext, alias=alias)
        """
        Primitive element used for the extension.

        >>> from sympy import QQ, sqrt
        >>> K = QQ.algebraic_field(sqrt(2), sqrt(3))
        >>> K.ext
        sqrt(2) + sqrt(3)
        """

        self.mod = self.ext.minpoly.rep
        """
        Minimal polynomial for the primitive element of the extension.

        >>> from sympy import QQ, sqrt
        >>> K = QQ.algebraic_field(sqrt(2))
        >>> K.mod
        DMP([1, 0, -2], QQ)
        """

        self.domain = self.dom = dom

        self.ngens = 1
        self.symbols = self.gens = (self.ext,)
        self.unit = self([dom(1), dom(0)])

        self.zero = self.dtype.zero(self.mod.to_list(), dom)
        self.one = self.dtype.one(self.mod.to_list(), dom)

        self._maximal_order = None
        self._discriminant = None
        self._nilradicals_mod_p = {}

    def new(self, element):
        return self.dtype(element, self.mod.to_list(), self.dom)

    def __str__(self):
        return str(self.dom) + '<' + str(self.ext) + '>'

    def __hash__(self):
        return hash((self.__class__.__name__, self.dtype, self.dom, self.ext))

    def __eq__(self, other):
        """Returns ``True`` if two domains are equivalent. """
        if isinstance(other, AlgebraicField):
            return self.dtype == other.dtype and self.ext == other.ext
        else:
            return NotImplemented

    def algebraic_field(self, *extension, alias=None):
        r"""Returns an algebraic field, i.e. `\mathbb{Q}(\alpha, \ldots)`. """
        return AlgebraicField(self.dom, *((self.ext,) + extension), alias=alias)

    def to_alg_num(self, a):
        """Convert ``a`` of ``dtype`` to an :py:class:`~.AlgebraicNumber`. """
        return self.ext.field_element(a)

    def to_sympy(self, a):
        """Convert ``a`` of ``dtype`` to a SymPy object. """
        # Precompute a converter to be reused:
        if not hasattr(self, '_converter'):
            self._converter = _make_converter(self)

        return self._converter(a)

    def from_sympy(self, a):
        """Convert SymPy's expression to ``dtype``. """
        try:
            return self([self.dom.from_sympy(a)])
        except CoercionFailed:
            pass

        from sympy.polys.numberfields import to_number_field

        try:
            return self(to_number_field(a, self.ext).native_coeffs())
        except (NotAlgebraic, IsomorphismFailed):
            raise CoercionFailed(
                "%s is not a valid algebraic number in %s" % (a, self))

    def from_ZZ(K1, a, K0):
        """Convert a Python ``int`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_ZZ_python(K1, a, K0):
        """Convert a Python ``int`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_QQ(K1, a, K0):
        """Convert a Python ``Fraction`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_QQ_python(K1, a, K0):
        """Convert a Python ``Fraction`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_ZZ_gmpy(K1, a, K0):
        """Convert a GMPY ``mpz`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_QQ_gmpy(K1, a, K0):
        """Convert a GMPY ``mpq`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_RealField(K1, a, K0):
        """Convert a mpmath ``mpf`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def get_ring(self):
        """Returns a ring associated with ``self``. """
        raise DomainError('there is no ring associated with %s' % self)

    def is_positive(self, a):
        """Returns True if ``a`` is positive. """
        return self.dom.is_positive(a.LC())

    def is_negative(self, a):
        """Returns True if ``a`` is negative. """
        return self.dom.is_negative(a.LC())

    def is_nonpositive(self, a):
        """Returns True if ``a`` is non-positive. """
        return self.dom.is_nonpositive(a.LC())

    def is_nonnegative(self, a):
        """Returns True if ``a`` is non-negative. """
        return self.dom.is_nonnegative(a.LC())

    def numer(self, a):
        """Returns numerator of ``a``. """
        return a

    def denom(self, a):
        """Returns denominator of ``a``. """
        return self.one

    def from_AlgebraicField(K1, a, K0):
        """Convert AlgebraicField element 'a' to another AlgebraicField """
        return K1.from_sympy(K0.to_sympy(a))

    def from_GaussianIntegerRing(K1, a, K0):
        """Convert a GaussianInteger element 'a' to ``dtype``. """
        return K1.from_sympy(K0.to_sympy(a))

    def from_GaussianRationalField(K1, a, K0):
        """Convert a GaussianRational element 'a' to ``dtype``. """
        return K1.from_sympy(K0.to_sympy(a))

    def _do_round_two(self):
        from sympy.polys.numberfields.basis import round_two
        ZK, dK = round_two(self, radicals=self._nilradicals_mod_p)
        self._maximal_order = ZK
        self._discriminant = dK

    def maximal_order(self):
        """
        Compute the maximal order, or ring of integers, of the field.

        Returns
        =======

        :py:class:`~sympy.polys.numberfields.modules.Submodule`.

        See Also
        ========

        integral_basis

        """
        if self._maximal_order is None:
            self._do_round_two()
        return self._maximal_order

    def integral_basis(self, fmt=None):
        r"""
        Get an integral basis for the field.

        Parameters
        ==========

        fmt : str, None, optional (default=None)
            If ``None``, return a list of :py:class:`~.ANP` instances.
            If ``"sympy"``, convert each element of the list to an
            :py:class:`~.Expr`, using ``self.to_sympy()``.
            If ``"alg"``, convert each element of the list to an
            :py:class:`~.AlgebraicNumber`, using ``self.to_alg_num()``.

        Examples
        ========

        >>> from sympy import QQ, AlgebraicNumber, sqrt
        >>> alpha = AlgebraicNumber(sqrt(5), alias='alpha')
        >>> k = QQ.algebraic_field(alpha)
        >>> B0 = k.integral_basis()
        >>> B1 = k.integral_basis(fmt='sympy')
        >>> B2 = k.integral_basis(fmt='alg')
        >>> print(B0[1])  # doctest: +SKIP
        ANP([mpq(1,2), mpq(1,2)], [mpq(1,1), mpq(0,1), mpq(-5,1)], QQ)
        >>> print(B1[1])
        1/2 + alpha/2
        >>> print(B2[1])
        alpha/2 + 1/2

        In the last two cases we get legible expressions, which print somewhat
        differently because of the different types involved:

        >>> print(type(B1[1]))
        <class 'sympy.core.add.Add'>
        >>> print(type(B2[1]))
        <class 'sympy.core.numbers.AlgebraicNumber'>

        See Also
        ========

        to_sympy
        to_alg_num
        maximal_order
        """
        ZK = self.maximal_order()
        M = ZK.QQ_matrix
        n = M.shape[1]
        B = [self.new(list(reversed(M[:, j].flat()))) for j in range(n)]
        if fmt == 'sympy':
            return [self.to_sympy(b) for b in B]
        elif fmt == 'alg':
            return [self.to_alg_num(b) for b in B]
        return B

    def discriminant(self):
        """Get the discriminant of the field."""
        if self._discriminant is None:
            self._do_round_two()
        return self._discriminant

    def primes_above(self, p):
        """Compute the prime ideals lying above a given rational prime *p*."""
        from sympy.polys.numberfields.primes import prime_decomp
        ZK = self.maximal_order()
        dK = self.discriminant()
        rad = self._nilradicals_mod_p.get(p)
        return prime_decomp(p, ZK=ZK, dK=dK, radical=rad)

    def galois_group(self, by_name=False, max_tries=30, randomize=False):
        """
        Compute the Galois group of the Galois closure of this field.

        Examples
        ========

        If the field is Galois, the order of the group will equal the degree
        of the field:

        >>> from sympy import QQ
        >>> from sympy.abc import x
        >>> k = QQ.alg_field_from_poly(x**4 + 1)
        >>> G, _ = k.galois_group()
        >>> G.order()
        4

        If the field is not Galois, then its Galois closure is a proper
        extension, and the order of the Galois group will be greater than the
        degree of the field:

        >>> k = QQ.alg_field_from_poly(x**4 - 2)
        >>> G, _ = k.galois_group()
        >>> G.order()
        8

        See Also
        ========

        sympy.polys.numberfields.galoisgroups.galois_group

        """
        return self.ext.minpoly_of_element().galois_group(
            by_name=by_name, max_tries=max_tries, randomize=randomize)


def _make_converter(K):
    """Construct the converter to convert back to Expr"""
    # Precompute the effect of converting to SymPy and expanding expressions
    # like (sqrt(2) + sqrt(3))**2. Asking Expr to do the expansion on every
    # conversion from K to Expr is slow. Here we compute the expansions for
    # each power of the generator and collect together the resulting algebraic
    # terms and the rational coefficients into a matrix.

    ext = K.ext.as_expr()
    todom = K.dom.from_sympy
    toexpr = K.dom.to_sympy

    if not ext.is_Add:
        powers = [ext**n for n in range(K.mod.degree())]
    else:
        # primitive_element generates a QQ-linear combination of lower degree
        # algebraic numbers to generate the higher degree extension e.g.
        # QQ<sqrt(2)+sqrt(3)> That means that we end up having high powers of low
        # degree algebraic numbers that can be reduced. Here we will use the
        # minimal polynomials of the algebraic numbers to reduce those powers
        # before converting to Expr.
        from sympy.polys.numberfields.minpoly import minpoly

        # Decompose ext as a linear combination of gens and make a symbol for
        # each gen.
        gens, coeffs = zip(*ext.as_coefficients_dict().items())
        syms = symbols(f'a:{len(gens)}', cls=Dummy)
        sym2gen = dict(zip(syms, gens))

        # Make a polynomial ring that can express ext and minpolys of all gens
        # in terms of syms.
        R = K.dom[syms]
        monoms = [R.ring.monomial_basis(i) for i in range(R.ngens)]
        ext_dict = {m: todom(c) for m, c in zip(monoms, coeffs)}
        ext_poly = R.ring.from_dict(ext_dict)
        minpolys = [R.from_sympy(minpoly(g, s)) for s, g in sym2gen.items()]

        # Compute all powers of ext_poly reduced modulo minpolys
        powers = [R.one, ext_poly]
        for n in range(2, K.mod.degree()):
            ext_poly_n = (powers[-1] * ext_poly).rem(minpolys)
            powers.append(ext_poly_n)

        # Convert the powers back to Expr. This will recombine some things like
        # sqrt(2)*sqrt(3) -> sqrt(6).
        powers = [p.as_expr().xreplace(sym2gen) for p in powers]

    # This also expands some rational powers
    powers = [p.expand() for p in powers]

    # Collect the rational coefficients and algebraic Expr that can
    # map the ANP coefficients into an expanded SymPy expression
    terms = [dict(t.as_coeff_Mul()[::-1] for t in Add.make_args(p)) for p in powers]
    algebraics = set().union(*terms)
    matrix = [[todom(t.get(a, S.Zero)) for t in terms] for a in algebraics]

    # Create a function to do the conversion efficiently:

    def converter(a):
        """Convert a to Expr using converter"""
        ai = a.to_list()[::-1]
        coeffs_dom = [sum(mij*aj for mij, aj in zip(mi, ai)) for mi in matrix]
        coeffs_sympy = [toexpr(c) for c in coeffs_dom]
        res = Add(*(Mul(c, a) for c, a in zip(coeffs_sympy, algebraics)))
        return res

    return converter

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\printing\rust.py ===
"""
Rust code printer

The `RustCodePrinter` converts SymPy expressions into Rust expressions.

A complete code generator, which uses `rust_code` extensively, can be found
in `sympy.utilities.codegen`. The `codegen` module can be used to generate
complete source code files.

"""

# Possible Improvement
#
# * make sure we follow Rust Style Guidelines_
# * make use of pattern matching
# * better support for reference
# * generate generic code and use trait to make sure they have specific methods
# * use crates_ to get more math support
#     - num_
#         + BigInt_, BigUint_
#         + Complex_
#         + Rational64_, Rational32_, BigRational_
#
# .. _crates: https://crates.io/
# .. _Guidelines: https://github.com/rust-lang/rust/tree/master/src/doc/style
# .. _num: http://rust-num.github.io/num/num/
# .. _BigInt: http://rust-num.github.io/num/num/bigint/struct.BigInt.html
# .. _BigUint: http://rust-num.github.io/num/num/bigint/struct.BigUint.html
# .. _Complex: http://rust-num.github.io/num/num/complex/struct.Complex.html
# .. _Rational32: http://rust-num.github.io/num/num/rational/type.Rational32.html
# .. _Rational64: http://rust-num.github.io/num/num/rational/type.Rational64.html
# .. _BigRational: http://rust-num.github.io/num/num/rational/type.BigRational.html

from __future__ import annotations
from functools import reduce
import operator
from typing import Any

from sympy.codegen.ast import (
    float32, float64, int32,
    real, integer,  bool_
)
from sympy.core import S, Rational, Float, Lambda
from sympy.core.expr import Expr
from sympy.core.numbers import equal_valued
from sympy.functions.elementary.integers import ceiling, floor
from sympy.printing.codeprinter import CodePrinter
from sympy.printing.precedence import PRECEDENCE

# Rust's methods for integer and float can be found at here :
#
# * `Rust - Primitive Type f64 <https://doc.rust-lang.org/std/primitive.f64.html>`_
# * `Rust - Primitive Type i64 <https://doc.rust-lang.org/std/primitive.i64.html>`_
#
# Function Style :
#
# 1. args[0].func(args[1:]), method with arguments
# 2. args[0].func(), method without arguments
# 3. args[1].func(), method without arguments (e.g. (e, x) => x.exp())
# 4. func(args), function with arguments

# dictionary mapping SymPy function to (argument_conditions, Rust_function).
# Used in RustCodePrinter._print_Function(self)

class float_floor(floor):
    """
    Same as `sympy.floor`, but mimics the Rust behavior of returning a float rather than an integer
    """
    def _eval_is_integer(self):
        return False

class float_ceiling(ceiling):
    """
    Same as `sympy.ceiling`, but mimics the Rust behavior of returning a float rather than an integer
    """
    def _eval_is_integer(self):
        return False


function_overrides = {
    "floor": (floor, float_floor),
    "ceiling": (ceiling, float_ceiling),
}

# f64 method in Rust
known_functions = {
    # "": "is_nan",
    # "": "is_infinite",
    # "": "is_finite",
    # "": "is_normal",
    # "": "classify",
    "float_floor": "floor",
    "float_ceiling": "ceil",
    # "": "round",
    # "": "trunc",
    # "": "fract",
    "Abs": "abs",
    # "": "signum",
    # "": "is_sign_positive",
    # "": "is_sign_negative",
    # "": "mul_add",
    "Pow": [(lambda base, exp: equal_valued(exp, -1), "recip", 2),   # 1.0/x
            (lambda base, exp: equal_valued(exp, 0.5), "sqrt", 2),   # x ** 0.5
            (lambda base, exp: equal_valued(exp, -0.5), "sqrt().recip", 2),   # 1/(x ** 0.5)
            (lambda base, exp: exp == Rational(1, 3), "cbrt", 2),    # x ** (1/3)
            (lambda base, exp: equal_valued(base, 2), "exp2", 3),    # 2 ** x
            (lambda base, exp: exp.is_integer, "powi", 1),           # x ** y, for i32
            (lambda base, exp: not exp.is_integer, "powf", 1)],      # x ** y, for f64
    "exp": [(lambda exp: True, "exp", 2)],   # e ** x
    "log": "ln",
    # "": "log",          # number.log(base)
    # "": "log2",
    # "": "log10",
    # "": "to_degrees",
    # "": "to_radians",
    "Max": "max",
    "Min": "min",
    # "": "hypot",        # (x**2 + y**2) ** 0.5
    "sin": "sin",
    "cos": "cos",
    "tan": "tan",
    "asin": "asin",
    "acos": "acos",
    "atan": "atan",
    "atan2": "atan2",
    # "": "sin_cos",
    # "": "exp_m1",       # e ** x - 1
    # "": "ln_1p",        # ln(1 + x)
    "sinh": "sinh",
    "cosh": "cosh",
    "tanh": "tanh",
    "asinh": "asinh",
    "acosh": "acosh",
    "atanh": "atanh",
    "sqrt": "sqrt",  # To enable automatic rewrites
}

# i64 method in Rust
# known_functions_i64 = {
#     "": "min_value",
#     "": "max_value",
#     "": "from_str_radix",
#     "": "count_ones",
#     "": "count_zeros",
#     "": "leading_zeros",
#     "": "trainling_zeros",
#     "": "rotate_left",
#     "": "rotate_right",
#     "": "swap_bytes",
#     "": "from_be",
#     "": "from_le",
#     "": "to_be",    # to big endian
#     "": "to_le",    # to little endian
#     "": "checked_add",
#     "": "checked_sub",
#     "": "checked_mul",
#     "": "checked_div",
#     "": "checked_rem",
#     "": "checked_neg",
#     "": "checked_shl",
#     "": "checked_shr",
#     "": "checked_abs",
#     "": "saturating_add",
#     "": "saturating_sub",
#     "": "saturating_mul",
#     "": "wrapping_add",
#     "": "wrapping_sub",
#     "": "wrapping_mul",
#     "": "wrapping_div",
#     "": "wrapping_rem",
#     "": "wrapping_neg",
#     "": "wrapping_shl",
#     "": "wrapping_shr",
#     "": "wrapping_abs",
#     "": "overflowing_add",
#     "": "overflowing_sub",
#     "": "overflowing_mul",
#     "": "overflowing_div",
#     "": "overflowing_rem",
#     "": "overflowing_neg",
#     "": "overflowing_shl",
#     "": "overflowing_shr",
#     "": "overflowing_abs",
#     "Pow": "pow",
#     "Abs": "abs",
#     "sign": "signum",
#     "": "is_positive",
#     "": "is_negnative",
# }

# These are the core reserved words in the Rust language. Taken from:
# https://doc.rust-lang.org/reference/keywords.html

reserved_words = ['abstract',
                  'as',
                  'async',
                  'await',
                  'become',
                  'box',
                  'break',
                  'const',
                  'continue',
                  'crate',
                  'do',
                  'dyn',
                  'else',
                  'enum',
                  'extern',
                  'false',
                  'final',
                  'fn',
                  'for',
                  'gen',
                  'if',
                  'impl',
                  'in',
                  'let',
                  'loop',
                  'macro',
                  'match',
                  'mod',
                  'move',
                  'mut',
                  'override',
                  'priv',
                  'pub',
                  'ref',
                  'return',
                  'Self',
                  'self',
                  'static',
                  'struct',
                  'super',
                  'trait',
                  'true',
                  'try',
                  'type',
                  'typeof',
                  'unsafe',
                  'unsized',
                  'use',
                  'virtual',
                  'where',
                  'while',
                  'yield']


class TypeCast(Expr):
    """
    The type casting operator of the Rust language.
    """

    def __init__(self, expr, type_) -> None:
        super().__init__()
        self.explicit = expr.is_integer and type_ is not integer
        self._assumptions = expr._assumptions
        if self.explicit:
            setattr(self, 'precedence', PRECEDENCE["Func"] + 10)

    @property
    def expr(self):
        return self.args[0]

    @property
    def type_(self):
        return self.args[1]

    def sort_key(self, order=None):
        return self.args[0].sort_key(order=order)


class RustCodePrinter(CodePrinter):
    """A printer to convert SymPy expressions to strings of Rust code"""
    printmethod = "_rust_code"
    language = "Rust"

    type_aliases = {
        integer: int32,
        real: float64,
    }

    type_mappings = {
        int32: 'i32',
        float32: 'f32',
        float64: 'f64',
        bool_: 'bool'
    }

    _default_settings: dict[str, Any] = dict(CodePrinter._default_settings, **{
        'precision': 17,
        'user_functions': {},
        'contract': True,
        'dereference': set(),
    })

    def __init__(self, settings={}):
        CodePrinter.__init__(self, settings)
        self.known_functions = dict(known_functions)
        userfuncs = settings.get('user_functions', {})
        self.known_functions.update(userfuncs)
        self._dereference = set(settings.get('dereference', []))
        self.reserved_words = set(reserved_words)
        self.function_overrides = function_overrides

    def _rate_index_position(self, p):
        return p*5

    def _get_statement(self, codestring):
        return "%s;" % codestring

    def _get_comment(self, text):
        return "// %s" % text

    def _declare_number_const(self, name, value):
        type_ = self.type_mappings[self.type_aliases[real]]
        return "const %s: %s = %s;" % (name, type_, value)

    def _format_code(self, lines):
        return self.indent_code(lines)

    def _traverse_matrix_indices(self, mat):
        rows, cols = mat.shape
        return ((i, j) for i in range(rows) for j in range(cols))

    def _get_loop_opening_ending(self, indices):
        open_lines = []
        close_lines = []
        loopstart = "for %(var)s in %(start)s..%(end)s {"
        for i in indices:
            # Rust arrays start at 0 and end at dimension-1
            open_lines.append(loopstart % {
                'var': self._print(i),
                'start': self._print(i.lower),
                'end': self._print(i.upper + 1)})
            close_lines.append("}")
        return open_lines, close_lines

    def _print_caller_var(self, expr):
        if len(expr.args) > 1:
            # for something like `sin(x + y + z)`,
            # make sure we can get '(x + y + z).sin()'
            # instead of 'x + y + z.sin()'
            return '(' + self._print(expr) + ')'
        elif expr.is_number:
            return self._print(expr, _type=True)
        else:
            return self._print(expr)

    def _print_Function(self, expr):
        """
        basic function for printing `Function`

        Function Style :

        1. args[0].func(args[1:]), method with arguments
        2. args[0].func(), method without arguments
        3. args[1].func(), method without arguments (e.g. (e, x) => x.exp())
        4. func(args), function with arguments
        """

        if expr.func.__name__ in self.known_functions:
            cond_func = self.known_functions[expr.func.__name__]
            func = None
            style = 1
            if isinstance(cond_func, str):
                func = cond_func
            else:
                for cond, func, style in cond_func:
                    if cond(*expr.args):
                        break
            if func is not None:
                if style == 1:
                    ret = "%(var)s.%(method)s(%(args)s)" % {
                        'var': self._print_caller_var(expr.args[0]),
                        'method': func,
                        'args': self.stringify(expr.args[1:], ", ") if len(expr.args) > 1 else ''
                    }
                elif style == 2:
                    ret = "%(var)s.%(method)s()" % {
                        'var': self._print_caller_var(expr.args[0]),
                        'method': func,
                    }
                elif style == 3:
                    ret = "%(var)s.%(method)s()" % {
                        'var': self._print_caller_var(expr.args[1]),
                        'method': func,
                    }
                else:
                    ret = "%(func)s(%(args)s)" % {
                        'func': func,
                        'args': self.stringify(expr.args, ", "),
                    }
                return ret
        elif hasattr(expr, '_imp_') and isinstance(expr._imp_, Lambda):
            # inlined function
            return self._print(expr._imp_(*expr.args))
        else:
            return self._print_not_supported(expr)

    def _print_Mul(self, expr):
        contains_floats = any(arg.is_real and not arg.is_integer for arg in expr.args)
        if contains_floats:
            expr = reduce(operator.mul,(self._cast_to_float(arg) if arg != -1 else arg for arg in expr.args))

        return super()._print_Mul(expr)

    def _print_Add(self, expr, order=None):
        contains_floats = any(arg.is_real and not arg.is_integer for arg in expr.args)
        if contains_floats:
            expr = reduce(operator.add, (self._cast_to_float(arg) for arg in expr.args))

        return super()._print_Add(expr, order)

    def _print_Pow(self, expr):
        if expr.base.is_integer and not expr.exp.is_integer:
            expr = type(expr)(Float(expr.base), expr.exp)
            return self._print(expr)
        return self._print_Function(expr)

    def _print_TypeCast(self, expr):
        if not expr.explicit:
            return self._print(expr.expr)
        else:
            return self._print(expr.expr) + ' as %s' % self.type_mappings[self.type_aliases[expr.type_]]

    def _print_Float(self, expr, _type=False):
        ret = super()._print_Float(expr)
        if _type:
            return ret + '_%s' % self.type_mappings[self.type_aliases[real]]
        else:
            return ret

    def _print_Integer(self, expr, _type=False):
        ret = super()._print_Integer(expr)
        if _type:
            return ret + '_%s' % self.type_mappings[self.type_aliases[integer]]
        else:
            return ret

    def _print_Rational(self, expr):
        p, q = int(expr.p), int(expr.q)
        float_suffix = self.type_mappings[self.type_aliases[real]]
        return '%d_%s/%d.0' % (p, float_suffix, q)

    def _print_Relational(self, expr):
        if (expr.lhs.is_integer and not expr.rhs.is_integer) or (expr.rhs.is_integer and not expr.lhs.is_integer):
            lhs = self._cast_to_float(expr.lhs)
            rhs = self._cast_to_float(expr.rhs)
        else:
            lhs = expr.lhs
            rhs = expr.rhs
        lhs_code = self._print(lhs)
        rhs_code = self._print(rhs)
        op = expr.rel_op
        return "{} {} {}".format(lhs_code, op, rhs_code)

    def _print_Indexed(self, expr):
        # calculate index for 1d array
        dims = expr.shape
        elem = S.Zero
        offset = S.One
        for i in reversed(range(expr.rank)):
            elem += expr.indices[i]*offset
            offset *= dims[i]
        return "%s[%s]" % (self._print(expr.base.label), self._print(elem))

    def _print_Idx(self, expr):
        return expr.label.name

    def _print_Dummy(self, expr):
        return expr.name

    def _print_Exp1(self, expr, _type=False):
        return "E"

    def _print_Pi(self, expr, _type=False):
        return 'PI'

    def _print_Infinity(self, expr, _type=False):
        return 'INFINITY'

    def _print_NegativeInfinity(self, expr, _type=False):
        return 'NEG_INFINITY'

    def _print_BooleanTrue(self, expr, _type=False):
        return "true"

    def _print_BooleanFalse(self, expr, _type=False):
        return "false"

    def _print_bool(self, expr, _type=False):
        return str(expr).lower()

    def _print_NaN(self, expr, _type=False):
        return "NAN"

    def _print_Piecewise(self, expr):
        if expr.args[-1].cond != True:
            # We need the last conditional to be a True, otherwise the resulting
            # function may not return a result.
            raise ValueError("All Piecewise expressions must contain an "
                             "(expr, True) statement to be used as a default "
                             "condition. Without one, the generated "
                             "expression may not evaluate to anything under "
                             "some condition.")
        lines = []

        for i, (e, c) in enumerate(expr.args):
            if i == 0:
                lines.append("if (%s) {" % self._print(c))
            elif i == len(expr.args) - 1 and c == True:
                lines[-1] += " else {"
            else:
                lines[-1] += " else if (%s) {" % self._print(c)
            code0 = self._print(e)
            lines.append(code0)
            lines.append("}")

        if self._settings['inline']:
            return " ".join(lines)
        else:
            return "\n".join(lines)

    def _print_ITE(self, expr):
        from sympy.functions import Piecewise
        return self._print(expr.rewrite(Piecewise, deep=False))

    def _print_MatrixBase(self, A):
        if A.cols == 1:
            return "[%s]" % ", ".join(self._print(a) for a in A)
        else:
            raise ValueError("Full Matrix Support in Rust need Crates (https://crates.io/keywords/matrix).")

    def _print_SparseRepMatrix(self, mat):
        # do not allow sparse matrices to be made dense
        return self._print_not_supported(mat)

    def _print_MatrixElement(self, expr):
        return "%s[%s]" % (expr.parent,
                           expr.j + expr.i*expr.parent.shape[1])

    def _print_Symbol(self, expr):

        name = super()._print_Symbol(expr)

        if expr in self._dereference:
            return '(*%s)' % name
        else:
            return name

    def _print_Assignment(self, expr):
        from sympy.tensor.indexed import IndexedBase
        lhs = expr.lhs
        rhs = expr.rhs
        if self._settings["contract"] and (lhs.has(IndexedBase) or
                rhs.has(IndexedBase)):
            # Here we check if there is looping to be done, and if so
            # print the required loops.
            return self._doprint_loops(rhs, lhs)
        else:
            lhs_code = self._print(lhs)
            rhs_code = self._print(rhs)
            return self._get_statement("%s = %s" % (lhs_code, rhs_code))

    def _print_sign(self, expr):
        arg = self._print(expr.args[0])
        return "(if (%s == 0.0) { 0.0 } else { (%s).signum() })" % (arg, arg)

    def _cast_to_float(self, expr):
        if not expr.is_number:
            return TypeCast(expr, real)
        elif expr.is_integer:
            return Float(expr)
        return expr

    def _can_print(self, name):
        """ Check if function ``name`` is either a known function or has its own
            printing method. Used to check if rewriting is possible."""

        # since the whole point of function_overrides is to enable proper printing,
        # we presume they all are printable

        return name in self.known_functions or name in function_overrides or getattr(self, '_print_{}'.format(name), False)

    def _collect_functions(self, expr):
        functions = set()
        if isinstance(expr, Expr):
            if expr.is_Function:
                functions.add(expr.func)
            for arg in expr.args:
                functions = functions.union(self._collect_functions(arg))
        return functions

    def _rewrite_known_functions(self, expr):
        if not isinstance(expr, Expr):
            return expr

        expression_functions = self._collect_functions(expr)
        rewriteable_functions = {
            name: (target_f, required_fs)
            for name, (target_f, required_fs) in self._rewriteable_functions.items()
            if self._can_print(target_f)
            and all(self._can_print(f) for f in required_fs)
        }
        for func in expression_functions:
            target_f, _ = rewriteable_functions.get(func.__name__, (None, None))
            if target_f:
                expr = expr.rewrite(target_f)
        return expr

    def indent_code(self, code):
        """Accepts a string of code or a list of code lines"""

        if isinstance(code, str):
            code_lines = self.indent_code(code.splitlines(True))
            return ''.join(code_lines)

        tab = "    "
        inc_token = ('{', '(', '{\n', '(\n')
        dec_token = ('}', ')')

        code = [ line.lstrip(' \t') for line in code ]

        increase = [ int(any(map(line.endswith, inc_token))) for line in code ]
        decrease = [ int(any(map(line.startswith, dec_token)))
                     for line in code ]

        pretty = []
        level = 0
        for n, line in enumerate(code):
            if line in ('', '\n'):
                pretty.append(line)
                continue
            level -= decrease[n]
            pretty.append("%s%s" % (tab*level, line))
            level += increase[n]
        return pretty

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\testing\_memory_streams.py ===
from __future__ import annotations

import operator
from collections.abc import Awaitable, Callable
from typing import TYPE_CHECKING, TypeVar

from .. import _core, _util
from .._highlevel_generic import StapledStream
from ..abc import ReceiveStream, SendStream

if TYPE_CHECKING:
    from typing_extensions import TypeAlias


AsyncHook: TypeAlias = Callable[[], Awaitable[object]]
# Would be nice to exclude awaitable here, but currently not possible.
SyncHook: TypeAlias = Callable[[], object]
SendStreamT = TypeVar("SendStreamT", bound=SendStream)
ReceiveStreamT = TypeVar("ReceiveStreamT", bound=ReceiveStream)


################################################################
# In-memory streams - Unbounded buffer version
################################################################


class _UnboundedByteQueue:
    def __init__(self) -> None:
        self._data = bytearray()
        self._closed = False
        self._lot = _core.ParkingLot()
        self._fetch_lock = _util.ConflictDetector(
            "another task is already fetching data",
        )

    # This object treats "close" as being like closing the send side of a
    # channel: so after close(), calling put() raises ClosedResourceError, and
    # calling the get() variants drains the buffer and then returns an empty
    # bytearray.
    def close(self) -> None:
        self._closed = True
        self._lot.unpark_all()

    def close_and_wipe(self) -> None:
        self._data = bytearray()
        self.close()

    def put(self, data: bytes | bytearray | memoryview) -> None:
        if self._closed:
            raise _core.ClosedResourceError("virtual connection closed")
        self._data += data
        self._lot.unpark_all()

    def _check_max_bytes(self, max_bytes: int | None) -> None:
        if max_bytes is None:
            return
        max_bytes = operator.index(max_bytes)
        if max_bytes < 1:
            raise ValueError("max_bytes must be >= 1")

    def _get_impl(self, max_bytes: int | None) -> bytearray:
        assert self._closed or self._data
        if max_bytes is None:
            max_bytes = len(self._data)
        if self._data:
            chunk = self._data[:max_bytes]
            del self._data[:max_bytes]
            assert chunk
            return chunk
        else:
            return bytearray()

    def get_nowait(self, max_bytes: int | None = None) -> bytearray:
        with self._fetch_lock:
            self._check_max_bytes(max_bytes)
            if not self._closed and not self._data:
                raise _core.WouldBlock
            return self._get_impl(max_bytes)

    async def get(self, max_bytes: int | None = None) -> bytearray:
        with self._fetch_lock:
            self._check_max_bytes(max_bytes)
            if not self._closed and not self._data:
                await self._lot.park()
            else:
                await _core.checkpoint()
            return self._get_impl(max_bytes)


@_util.final
class MemorySendStream(SendStream):
    """An in-memory :class:`~trio.abc.SendStream`.

    Args:
      send_all_hook: An async function, or None. Called from
          :meth:`send_all`. Can do whatever you like.
      wait_send_all_might_not_block_hook: An async function, or None. Called
          from :meth:`wait_send_all_might_not_block`. Can do whatever you
          like.
      close_hook: A synchronous function, or None. Called from :meth:`close`
          and :meth:`aclose`. Can do whatever you like.

    .. attribute:: send_all_hook
                   wait_send_all_might_not_block_hook
                   close_hook

       All of these hooks are also exposed as attributes on the object, and
       you can change them at any time.

    """

    def __init__(
        self,
        send_all_hook: AsyncHook | None = None,
        wait_send_all_might_not_block_hook: AsyncHook | None = None,
        close_hook: SyncHook | None = None,
    ) -> None:
        self._conflict_detector = _util.ConflictDetector(
            "another task is using this stream",
        )
        self._outgoing = _UnboundedByteQueue()
        self.send_all_hook = send_all_hook
        self.wait_send_all_might_not_block_hook = wait_send_all_might_not_block_hook
        self.close_hook = close_hook

    async def send_all(self, data: bytes | bytearray | memoryview) -> None:
        """Places the given data into the object's internal buffer, and then
        calls the :attr:`send_all_hook` (if any).

        """
        # Execute two checkpoints so we have more of a chance to detect
        # buggy user code that calls this twice at the same time.
        with self._conflict_detector:
            await _core.checkpoint()
            await _core.checkpoint()
            self._outgoing.put(data)
            if self.send_all_hook is not None:
                await self.send_all_hook()

    async def wait_send_all_might_not_block(self) -> None:
        """Calls the :attr:`wait_send_all_might_not_block_hook` (if any), and
        then returns immediately.

        """
        # Execute two checkpoints so that we have more of a chance to detect
        # buggy user code that calls this twice at the same time.
        with self._conflict_detector:
            await _core.checkpoint()
            await _core.checkpoint()
            # check for being closed:
            self._outgoing.put(b"")
            if self.wait_send_all_might_not_block_hook is not None:
                await self.wait_send_all_might_not_block_hook()

    def close(self) -> None:
        """Marks this stream as closed, and then calls the :attr:`close_hook`
        (if any).

        """
        # XXX should this cancel any pending calls to the send_all_hook and
        # wait_send_all_might_not_block_hook? Those are the only places where
        # send_all and wait_send_all_might_not_block can be blocked.
        #
        # The way we set things up, send_all_hook is memory_stream_pump, and
        # wait_send_all_might_not_block_hook is unset. memory_stream_pump is
        # synchronous. So normally, send_all and wait_send_all_might_not_block
        # cannot block at all.
        self._outgoing.close()
        if self.close_hook is not None:
            self.close_hook()

    async def aclose(self) -> None:
        """Same as :meth:`close`, but async."""
        self.close()
        await _core.checkpoint()

    async def get_data(self, max_bytes: int | None = None) -> bytearray:
        """Retrieves data from the internal buffer, blocking if necessary.

        Args:
          max_bytes (int or None): The maximum amount of data to
              retrieve. None (the default) means to retrieve all the data
              that's present (but still blocks until at least one byte is
              available).

        Returns:
          If this stream has been closed, an empty bytearray. Otherwise, the
          requested data.

        """
        return await self._outgoing.get(max_bytes)

    def get_data_nowait(self, max_bytes: int | None = None) -> bytearray:
        """Retrieves data from the internal buffer, but doesn't block.

        See :meth:`get_data` for details.

        Raises:
          trio.WouldBlock: if no data is available to retrieve.

        """
        return self._outgoing.get_nowait(max_bytes)


@_util.final
class MemoryReceiveStream(ReceiveStream):
    """An in-memory :class:`~trio.abc.ReceiveStream`.

    Args:
      receive_some_hook: An async function, or None. Called from
          :meth:`receive_some`. Can do whatever you like.
      close_hook: A synchronous function, or None. Called from :meth:`close`
          and :meth:`aclose`. Can do whatever you like.

    .. attribute:: receive_some_hook
                   close_hook

       Both hooks are also exposed as attributes on the object, and you can
       change them at any time.

    """

    def __init__(
        self,
        receive_some_hook: AsyncHook | None = None,
        close_hook: SyncHook | None = None,
    ) -> None:
        self._conflict_detector = _util.ConflictDetector(
            "another task is using this stream",
        )
        self._incoming = _UnboundedByteQueue()
        self._closed = False
        self.receive_some_hook = receive_some_hook
        self.close_hook = close_hook

    async def receive_some(self, max_bytes: int | None = None) -> bytearray:
        """Calls the :attr:`receive_some_hook` (if any), and then retrieves
        data from the internal buffer, blocking if necessary.

        """
        # Execute two checkpoints so we have more of a chance to detect
        # buggy user code that calls this twice at the same time.
        with self._conflict_detector:
            await _core.checkpoint()
            await _core.checkpoint()
            if self._closed:
                raise _core.ClosedResourceError
            if self.receive_some_hook is not None:
                await self.receive_some_hook()
            # self._incoming's closure state tracks whether we got an EOF.
            # self._closed tracks whether we, ourselves, are closed.
            # self.close() sends an EOF to wake us up and sets self._closed,
            # so after we wake up we have to check self._closed again.
            data = await self._incoming.get(max_bytes)
            if self._closed:
                raise _core.ClosedResourceError
            return data

    def close(self) -> None:
        """Discards any pending data from the internal buffer, and marks this
        stream as closed.

        """
        self._closed = True
        self._incoming.close_and_wipe()
        if self.close_hook is not None:
            self.close_hook()

    async def aclose(self) -> None:
        """Same as :meth:`close`, but async."""
        self.close()
        await _core.checkpoint()

    def put_data(self, data: bytes | bytearray | memoryview) -> None:
        """Appends the given data to the internal buffer."""
        self._incoming.put(data)

    def put_eof(self) -> None:
        """Adds an end-of-file marker to the internal buffer."""
        self._incoming.close()


# TODO: investigate why this is necessary for the docs
MemorySendStream.__module__ = MemorySendStream.__module__.replace(
    "._memory_streams", ""
)
MemoryReceiveStream.__module__ = MemoryReceiveStream.__module__.replace(
    "._memory_streams", ""
)


def memory_stream_pump(
    memory_send_stream: MemorySendStream,
    memory_receive_stream: MemoryReceiveStream,
    *,
    max_bytes: int | None = None,
) -> bool:
    """Take data out of the given :class:`MemorySendStream`'s internal buffer,
    and put it into the given :class:`MemoryReceiveStream`'s internal buffer.

    Args:
      memory_send_stream (MemorySendStream): The stream to get data from.
      memory_receive_stream (MemoryReceiveStream): The stream to put data into.
      max_bytes (int or None): The maximum amount of data to transfer in this
          call, or None to transfer all available data.

    Returns:
      True if it successfully transferred some data, or False if there was no
      data to transfer.

    This is used to implement :func:`memory_stream_one_way_pair` and
    :func:`memory_stream_pair`; see the latter's docstring for an example
    of how you might use it yourself.

    """
    try:
        data = memory_send_stream.get_data_nowait(max_bytes)
    except _core.WouldBlock:
        return False
    try:
        if not data:
            memory_receive_stream.put_eof()
        else:
            memory_receive_stream.put_data(data)
    except _core.ClosedResourceError:
        raise _core.BrokenResourceError("MemoryReceiveStream was closed") from None
    return True


def memory_stream_one_way_pair() -> tuple[MemorySendStream, MemoryReceiveStream]:
    """Create a connected, pure-Python, unidirectional stream with infinite
    buffering and flexible configuration options.

    You can think of this as being a no-operating-system-involved
    Trio-streamsified version of :func:`os.pipe` (except that :func:`os.pipe`
    returns the streams in the wrong order – we follow the superior convention
    that data flows from left to right).

    Returns:
      A tuple (:class:`MemorySendStream`, :class:`MemoryReceiveStream`), where
      the :class:`MemorySendStream` has its hooks set up so that it calls
      :func:`memory_stream_pump` from its
      :attr:`~MemorySendStream.send_all_hook` and
      :attr:`~MemorySendStream.close_hook`.

    The end result is that data automatically flows from the
    :class:`MemorySendStream` to the :class:`MemoryReceiveStream`. But you're
    also free to rearrange things however you like. For example, you can
    temporarily set the :attr:`~MemorySendStream.send_all_hook` to None if you
    want to simulate a stall in data transmission. Or see
    :func:`memory_stream_pair` for a more elaborate example.

    """
    send_stream = MemorySendStream()
    recv_stream = MemoryReceiveStream()

    def pump_from_send_stream_to_recv_stream() -> None:
        memory_stream_pump(send_stream, recv_stream)

    # await not used
    async def async_pump_from_send_stream_to_recv_stream() -> None:  # noqa: RUF029
        pump_from_send_stream_to_recv_stream()

    send_stream.send_all_hook = async_pump_from_send_stream_to_recv_stream
    send_stream.close_hook = pump_from_send_stream_to_recv_stream
    return send_stream, recv_stream


def _make_stapled_pair(
    one_way_pair: Callable[[], tuple[SendStreamT, ReceiveStreamT]],
) -> tuple[
    StapledStream[SendStreamT, ReceiveStreamT],
    StapledStream[SendStreamT, ReceiveStreamT],
]:
    pipe1_send, pipe1_recv = one_way_pair()
    pipe2_send, pipe2_recv = one_way_pair()
    stream1 = StapledStream(pipe1_send, pipe2_recv)
    stream2 = StapledStream(pipe2_send, pipe1_recv)
    return stream1, stream2


def memory_stream_pair() -> tuple[
    StapledStream[MemorySendStream, MemoryReceiveStream],
    StapledStream[MemorySendStream, MemoryReceiveStream],
]:
    """Create a connected, pure-Python, bidirectional stream with infinite
    buffering and flexible configuration options.

    This is a convenience function that creates two one-way streams using
    :func:`memory_stream_one_way_pair`, and then uses
    :class:`~trio.StapledStream` to combine them into a single bidirectional
    stream.

    This is like a no-operating-system-involved, Trio-streamsified version of
    :func:`socket.socketpair`.

    Returns:
      A pair of :class:`~trio.StapledStream` objects that are connected so
      that data automatically flows from one to the other in both directions.

    After creating a stream pair, you can send data back and forth, which is
    enough for simple tests::

       left, right = memory_stream_pair()
       await left.send_all(b"123")
       assert await right.receive_some() == b"123"
       await right.send_all(b"456")
       assert await left.receive_some() == b"456"

    But if you read the docs for :class:`~trio.StapledStream` and
    :func:`memory_stream_one_way_pair`, you'll see that all the pieces
    involved in wiring this up are public APIs, so you can adjust to suit the
    requirements of your tests. For example, here's how to tweak a stream so
    that data flowing from left to right trickles in one byte at a time (but
    data flowing from right to left proceeds at full speed)::

        left, right = memory_stream_pair()
        async def trickle():
            # left is a StapledStream, and left.send_stream is a MemorySendStream
            # right is a StapledStream, and right.recv_stream is a MemoryReceiveStream
            while memory_stream_pump(left.send_stream, right.recv_stream, max_bytes=1):
                # Pause between each byte
                await trio.sleep(1)
        # Normally this send_all_hook calls memory_stream_pump directly without
        # passing in a max_bytes. We replace it with our custom version:
        left.send_stream.send_all_hook = trickle

    And here's a simple test using our modified stream objects::

        async def sender():
            await left.send_all(b"12345")
            await left.send_eof()

        async def receiver():
            async for data in right:
                print(data)

        async with trio.open_nursery() as nursery:
            nursery.start_soon(sender)
            nursery.start_soon(receiver)

    By default, this will print ``b"12345"`` and then immediately exit; with
    our trickle stream it instead sleeps 1 second, then prints ``b"1"``, then
    sleeps 1 second, then prints ``b"2"``, etc.

    Pro-tip: you can insert sleep calls (like in our example above) to
    manipulate the flow of data across tasks... and then use
    :class:`MockClock` and its :attr:`~MockClock.autojump_threshold`
    functionality to keep your test suite running quickly.

    If you want to stress test a protocol implementation, one nice trick is to
    use the :mod:`random` module (preferably with a fixed seed) to move random
    numbers of bytes at a time, and insert random sleeps in between them. You
    can also set up a custom :attr:`~MemoryReceiveStream.receive_some_hook` if
    you want to manipulate things on the receiving side, and not just the
    sending side.

    """
    return _make_stapled_pair(memory_stream_one_way_pair)


################################################################
# In-memory streams - Lockstep version
################################################################


class _LockstepByteQueue:
    def __init__(self) -> None:
        self._data = bytearray()
        self._sender_closed = False
        self._receiver_closed = False
        self._receiver_waiting = False
        self._waiters = _core.ParkingLot()
        self._send_conflict_detector = _util.ConflictDetector(
            "another task is already sending",
        )
        self._receive_conflict_detector = _util.ConflictDetector(
            "another task is already receiving",
        )

    def _something_happened(self) -> None:
        self._waiters.unpark_all()

    # Always wakes up when one side is closed, because everyone always reacts
    # to that.
    async def _wait_for(self, fn: Callable[[], bool]) -> None:
        while True:
            if fn():
                break
            if self._sender_closed or self._receiver_closed:
                break
            await self._waiters.park()
        await _core.checkpoint()

    def close_sender(self) -> None:
        self._sender_closed = True
        self._something_happened()

    def close_receiver(self) -> None:
        self._receiver_closed = True
        self._something_happened()

    async def send_all(self, data: bytes | bytearray | memoryview) -> None:
        with self._send_conflict_detector:
            if self._sender_closed:
                raise _core.ClosedResourceError
            if self._receiver_closed:
                raise _core.BrokenResourceError
            assert not self._data
            self._data += data
            self._something_happened()
            await self._wait_for(lambda: self._data == b"")
            if self._sender_closed:
                raise _core.ClosedResourceError
            if self._data and self._receiver_closed:
                raise _core.BrokenResourceError

    async def wait_send_all_might_not_block(self) -> None:
        with self._send_conflict_detector:
            if self._sender_closed:
                raise _core.ClosedResourceError
            if self._receiver_closed:
                await _core.checkpoint()
                return
            await self._wait_for(lambda: self._receiver_waiting)
            if self._sender_closed:
                raise _core.ClosedResourceError

    async def receive_some(self, max_bytes: int | None = None) -> bytes | bytearray:
        with self._receive_conflict_detector:
            # Argument validation
            if max_bytes is not None:
                max_bytes = operator.index(max_bytes)
                if max_bytes < 1:
                    raise ValueError("max_bytes must be >= 1")
            # State validation
            if self._receiver_closed:
                raise _core.ClosedResourceError
            # Wake wait_send_all_might_not_block and wait for data
            self._receiver_waiting = True
            self._something_happened()
            try:
                await self._wait_for(lambda: self._data != b"")
            finally:
                self._receiver_waiting = False
            if self._receiver_closed:
                raise _core.ClosedResourceError
            # Get data, possibly waking send_all
            if self._data:
                # Neat trick: if max_bytes is None, then obj[:max_bytes] is
                # the same as obj[:].
                got = self._data[:max_bytes]
                del self._data[:max_bytes]
                self._something_happened()
                return got
            else:
                assert self._sender_closed
                return b""


class _LockstepSendStream(SendStream):
    def __init__(self, lbq: _LockstepByteQueue) -> None:
        self._lbq = lbq

    def close(self) -> None:
        self._lbq.close_sender()

    async def aclose(self) -> None:
        self.close()
        await _core.checkpoint()

    async def send_all(self, data: bytes | bytearray | memoryview) -> None:
        await self._lbq.send_all(data)

    async def wait_send_all_might_not_block(self) -> None:
        await self._lbq.wait_send_all_might_not_block()


class _LockstepReceiveStream(ReceiveStream):
    def __init__(self, lbq: _LockstepByteQueue) -> None:
        self._lbq = lbq

    def close(self) -> None:
        self._lbq.close_receiver()

    async def aclose(self) -> None:
        self.close()
        await _core.checkpoint()

    async def receive_some(self, max_bytes: int | None = None) -> bytes | bytearray:
        return await self._lbq.receive_some(max_bytes)


def lockstep_stream_one_way_pair() -> tuple[SendStream, ReceiveStream]:
    """Create a connected, pure Python, unidirectional stream where data flows
    in lockstep.

    Returns:
      A tuple
      (:class:`~trio.abc.SendStream`, :class:`~trio.abc.ReceiveStream`).

    This stream has *absolutely no* buffering. Each call to
    :meth:`~trio.abc.SendStream.send_all` will block until all the given data
    has been returned by a call to
    :meth:`~trio.abc.ReceiveStream.receive_some`.

    This can be useful for testing flow control mechanisms in an extreme case,
    or for setting up "clogged" streams to use with
    :func:`check_one_way_stream` and friends.

    In addition to fulfilling the :class:`~trio.abc.SendStream` and
    :class:`~trio.abc.ReceiveStream` interfaces, the return objects
    also have a synchronous ``close`` method.

    """

    lbq = _LockstepByteQueue()
    return _LockstepSendStream(lbq), _LockstepReceiveStream(lbq)


def lockstep_stream_pair() -> tuple[
    StapledStream[SendStream, ReceiveStream],
    StapledStream[SendStream, ReceiveStream],
]:
    """Create a connected, pure-Python, bidirectional stream where data flows
    in lockstep.

    Returns:
      A tuple (:class:`~trio.StapledStream`, :class:`~trio.StapledStream`).

    This is a convenience function that creates two one-way streams using
    :func:`lockstep_stream_one_way_pair`, and then uses
    :class:`~trio.StapledStream` to combine them into a single bidirectional
    stream.

    """
    return _make_stapled_pair(lockstep_stream_one_way_pair)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\urllib3\poolmanager.py ===
from __future__ import annotations

import functools
import logging
import typing
import warnings
from types import TracebackType
from urllib.parse import urljoin

from ._collections import HTTPHeaderDict, RecentlyUsedContainer
from ._request_methods import RequestMethods
from .connection import ProxyConfig
from .connectionpool import HTTPConnectionPool, HTTPSConnectionPool, port_by_scheme
from .exceptions import (
    LocationValueError,
    MaxRetryError,
    ProxySchemeUnknown,
    URLSchemeUnknown,
)
from .response import BaseHTTPResponse
from .util.connection import _TYPE_SOCKET_OPTIONS
from .util.proxy import connection_requires_http_tunnel
from .util.retry import Retry
from .util.timeout import Timeout
from .util.url import Url, parse_url

if typing.TYPE_CHECKING:
    import ssl

    from typing_extensions import Self

__all__ = ["PoolManager", "ProxyManager", "proxy_from_url"]


log = logging.getLogger(__name__)

SSL_KEYWORDS = (
    "key_file",
    "cert_file",
    "cert_reqs",
    "ca_certs",
    "ca_cert_data",
    "ssl_version",
    "ssl_minimum_version",
    "ssl_maximum_version",
    "ca_cert_dir",
    "ssl_context",
    "key_password",
    "server_hostname",
)
# Default value for `blocksize` - a new parameter introduced to
# http.client.HTTPConnection & http.client.HTTPSConnection in Python 3.7
_DEFAULT_BLOCKSIZE = 16384


class PoolKey(typing.NamedTuple):
    """
    All known keyword arguments that could be provided to the pool manager, its
    pools, or the underlying connections.

    All custom key schemes should include the fields in this key at a minimum.
    """

    key_scheme: str
    key_host: str
    key_port: int | None
    key_timeout: Timeout | float | int | None
    key_retries: Retry | bool | int | None
    key_block: bool | None
    key_source_address: tuple[str, int] | None
    key_key_file: str | None
    key_key_password: str | None
    key_cert_file: str | None
    key_cert_reqs: str | None
    key_ca_certs: str | None
    key_ca_cert_data: str | bytes | None
    key_ssl_version: int | str | None
    key_ssl_minimum_version: ssl.TLSVersion | None
    key_ssl_maximum_version: ssl.TLSVersion | None
    key_ca_cert_dir: str | None
    key_ssl_context: ssl.SSLContext | None
    key_maxsize: int | None
    key_headers: frozenset[tuple[str, str]] | None
    key__proxy: Url | None
    key__proxy_headers: frozenset[tuple[str, str]] | None
    key__proxy_config: ProxyConfig | None
    key_socket_options: _TYPE_SOCKET_OPTIONS | None
    key__socks_options: frozenset[tuple[str, str]] | None
    key_assert_hostname: bool | str | None
    key_assert_fingerprint: str | None
    key_server_hostname: str | None
    key_blocksize: int | None


def _default_key_normalizer(
    key_class: type[PoolKey], request_context: dict[str, typing.Any]
) -> PoolKey:
    """
    Create a pool key out of a request context dictionary.

    According to RFC 3986, both the scheme and host are case-insensitive.
    Therefore, this function normalizes both before constructing the pool
    key for an HTTPS request. If you wish to change this behaviour, provide
    alternate callables to ``key_fn_by_scheme``.

    :param key_class:
        The class to use when constructing the key. This should be a namedtuple
        with the ``scheme`` and ``host`` keys at a minimum.
    :type  key_class: namedtuple
    :param request_context:
        A dictionary-like object that contain the context for a request.
    :type  request_context: dict

    :return: A namedtuple that can be used as a connection pool key.
    :rtype:  PoolKey
    """
    # Since we mutate the dictionary, make a copy first
    context = request_context.copy()
    context["scheme"] = context["scheme"].lower()
    context["host"] = context["host"].lower()

    # These are both dictionaries and need to be transformed into frozensets
    for key in ("headers", "_proxy_headers", "_socks_options"):
        if key in context and context[key] is not None:
            context[key] = frozenset(context[key].items())

    # The socket_options key may be a list and needs to be transformed into a
    # tuple.
    socket_opts = context.get("socket_options")
    if socket_opts is not None:
        context["socket_options"] = tuple(socket_opts)

    # Map the kwargs to the names in the namedtuple - this is necessary since
    # namedtuples can't have fields starting with '_'.
    for key in list(context.keys()):
        context["key_" + key] = context.pop(key)

    # Default to ``None`` for keys missing from the context
    for field in key_class._fields:
        if field not in context:
            context[field] = None

    # Default key_blocksize to _DEFAULT_BLOCKSIZE if missing from the context
    if context.get("key_blocksize") is None:
        context["key_blocksize"] = _DEFAULT_BLOCKSIZE

    return key_class(**context)


#: A dictionary that maps a scheme to a callable that creates a pool key.
#: This can be used to alter the way pool keys are constructed, if desired.
#: Each PoolManager makes a copy of this dictionary so they can be configured
#: globally here, or individually on the instance.
key_fn_by_scheme = {
    "http": functools.partial(_default_key_normalizer, PoolKey),
    "https": functools.partial(_default_key_normalizer, PoolKey),
}

pool_classes_by_scheme = {"http": HTTPConnectionPool, "https": HTTPSConnectionPool}


class PoolManager(RequestMethods):
    """
    Allows for arbitrary requests while transparently keeping track of
    necessary connection pools for you.

    :param num_pools:
        Number of connection pools to cache before discarding the least
        recently used pool.

    :param headers:
        Headers to include with all requests, unless other headers are given
        explicitly.

    :param \\**connection_pool_kw:
        Additional parameters are used to create fresh
        :class:`urllib3.connectionpool.ConnectionPool` instances.

    Example:

    .. code-block:: python

        import urllib3

        http = urllib3.PoolManager(num_pools=2)

        resp1 = http.request("GET", "https://google.com/")
        resp2 = http.request("GET", "https://google.com/mail")
        resp3 = http.request("GET", "https://yahoo.com/")

        print(len(http.pools))
        # 2

    """

    proxy: Url | None = None
    proxy_config: ProxyConfig | None = None

    def __init__(
        self,
        num_pools: int = 10,
        headers: typing.Mapping[str, str] | None = None,
        **connection_pool_kw: typing.Any,
    ) -> None:
        super().__init__(headers)
        self.connection_pool_kw = connection_pool_kw

        self.pools: RecentlyUsedContainer[PoolKey, HTTPConnectionPool]
        self.pools = RecentlyUsedContainer(num_pools)

        # Locally set the pool classes and keys so other PoolManagers can
        # override them.
        self.pool_classes_by_scheme = pool_classes_by_scheme
        self.key_fn_by_scheme = key_fn_by_scheme.copy()

    def __enter__(self) -> Self:
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: TracebackType | None,
    ) -> typing.Literal[False]:
        self.clear()
        # Return False to re-raise any potential exceptions
        return False

    def _new_pool(
        self,
        scheme: str,
        host: str,
        port: int,
        request_context: dict[str, typing.Any] | None = None,
    ) -> HTTPConnectionPool:
        """
        Create a new :class:`urllib3.connectionpool.ConnectionPool` based on host, port, scheme, and
        any additional pool keyword arguments.

        If ``request_context`` is provided, it is provided as keyword arguments
        to the pool class used. This method is used to actually create the
        connection pools handed out by :meth:`connection_from_url` and
        companion methods. It is intended to be overridden for customization.
        """
        pool_cls: type[HTTPConnectionPool] = self.pool_classes_by_scheme[scheme]
        if request_context is None:
            request_context = self.connection_pool_kw.copy()

        # Default blocksize to _DEFAULT_BLOCKSIZE if missing or explicitly
        # set to 'None' in the request_context.
        if request_context.get("blocksize") is None:
            request_context["blocksize"] = _DEFAULT_BLOCKSIZE

        # Although the context has everything necessary to create the pool,
        # this function has historically only used the scheme, host, and port
        # in the positional args. When an API change is acceptable these can
        # be removed.
        for key in ("scheme", "host", "port"):
            request_context.pop(key, None)

        if scheme == "http":
            for kw in SSL_KEYWORDS:
                request_context.pop(kw, None)

        return pool_cls(host, port, **request_context)

    def clear(self) -> None:
        """
        Empty our store of pools and direct them all to close.

        This will not affect in-flight connections, but they will not be
        re-used after completion.
        """
        self.pools.clear()

    def connection_from_host(
        self,
        host: str | None,
        port: int | None = None,
        scheme: str | None = "http",
        pool_kwargs: dict[str, typing.Any] | None = None,
    ) -> HTTPConnectionPool:
        """
        Get a :class:`urllib3.connectionpool.ConnectionPool` based on the host, port, and scheme.

        If ``port`` isn't given, it will be derived from the ``scheme`` using
        ``urllib3.connectionpool.port_by_scheme``. If ``pool_kwargs`` is
        provided, it is merged with the instance's ``connection_pool_kw``
        variable and used to create the new connection pool, if one is
        needed.
        """

        if not host:
            raise LocationValueError("No host specified.")

        request_context = self._merge_pool_kwargs(pool_kwargs)
        request_context["scheme"] = scheme or "http"
        if not port:
            port = port_by_scheme.get(request_context["scheme"].lower(), 80)
        request_context["port"] = port
        request_context["host"] = host

        return self.connection_from_context(request_context)

    def connection_from_context(
        self, request_context: dict[str, typing.Any]
    ) -> HTTPConnectionPool:
        """
        Get a :class:`urllib3.connectionpool.ConnectionPool` based on the request context.

        ``request_context`` must at least contain the ``scheme`` key and its
        value must be a key in ``key_fn_by_scheme`` instance variable.
        """
        if "strict" in request_context:
            warnings.warn(
                "The 'strict' parameter is no longer needed on Python 3+. "
                "This will raise an error in urllib3 v2.1.0.",
                DeprecationWarning,
            )
            request_context.pop("strict")

        scheme = request_context["scheme"].lower()
        pool_key_constructor = self.key_fn_by_scheme.get(scheme)
        if not pool_key_constructor:
            raise URLSchemeUnknown(scheme)
        pool_key = pool_key_constructor(request_context)

        return self.connection_from_pool_key(pool_key, request_context=request_context)

    def connection_from_pool_key(
        self, pool_key: PoolKey, request_context: dict[str, typing.Any]
    ) -> HTTPConnectionPool:
        """
        Get a :class:`urllib3.connectionpool.ConnectionPool` based on the provided pool key.

        ``pool_key`` should be a namedtuple that only contains immutable
        objects. At a minimum it must have the ``scheme``, ``host``, and
        ``port`` fields.
        """
        with self.pools.lock:
            # If the scheme, host, or port doesn't match existing open
            # connections, open a new ConnectionPool.
            pool = self.pools.get(pool_key)
            if pool:
                return pool

            # Make a fresh ConnectionPool of the desired type
            scheme = request_context["scheme"]
            host = request_context["host"]
            port = request_context["port"]
            pool = self._new_pool(scheme, host, port, request_context=request_context)
            self.pools[pool_key] = pool

        return pool

    def connection_from_url(
        self, url: str, pool_kwargs: dict[str, typing.Any] | None = None
    ) -> HTTPConnectionPool:
        """
        Similar to :func:`urllib3.connectionpool.connection_from_url`.

        If ``pool_kwargs`` is not provided and a new pool needs to be
        constructed, ``self.connection_pool_kw`` is used to initialize
        the :class:`urllib3.connectionpool.ConnectionPool`. If ``pool_kwargs``
        is provided, it is used instead. Note that if a new pool does not
        need to be created for the request, the provided ``pool_kwargs`` are
        not used.
        """
        u = parse_url(url)
        return self.connection_from_host(
            u.host, port=u.port, scheme=u.scheme, pool_kwargs=pool_kwargs
        )

    def _merge_pool_kwargs(
        self, override: dict[str, typing.Any] | None
    ) -> dict[str, typing.Any]:
        """
        Merge a dictionary of override values for self.connection_pool_kw.

        This does not modify self.connection_pool_kw and returns a new dict.
        Any keys in the override dictionary with a value of ``None`` are
        removed from the merged dictionary.
        """
        base_pool_kwargs = self.connection_pool_kw.copy()
        if override:
            for key, value in override.items():
                if value is None:
                    try:
                        del base_pool_kwargs[key]
                    except KeyError:
                        pass
                else:
                    base_pool_kwargs[key] = value
        return base_pool_kwargs

    def _proxy_requires_url_absolute_form(self, parsed_url: Url) -> bool:
        """
        Indicates if the proxy requires the complete destination URL in the
        request.  Normally this is only needed when not using an HTTP CONNECT
        tunnel.
        """
        if self.proxy is None:
            return False

        return not connection_requires_http_tunnel(
            self.proxy, self.proxy_config, parsed_url.scheme
        )

    def urlopen(  # type: ignore[override]
        self, method: str, url: str, redirect: bool = True, **kw: typing.Any
    ) -> BaseHTTPResponse:
        """
        Same as :meth:`urllib3.HTTPConnectionPool.urlopen`
        with custom cross-host redirect logic and only sends the request-uri
        portion of the ``url``.

        The given ``url`` parameter must be absolute, such that an appropriate
        :class:`urllib3.connectionpool.ConnectionPool` can be chosen for it.
        """
        u = parse_url(url)

        if u.scheme is None:
            warnings.warn(
                "URLs without a scheme (ie 'https://') are deprecated and will raise an error "
                "in a future version of urllib3. To avoid this DeprecationWarning ensure all URLs "
                "start with 'https://' or 'http://'. Read more in this issue: "
                "https://github.com/urllib3/urllib3/issues/2920",
                category=DeprecationWarning,
                stacklevel=2,
            )

        conn = self.connection_from_host(u.host, port=u.port, scheme=u.scheme)

        kw["assert_same_host"] = False
        kw["redirect"] = False

        if "headers" not in kw:
            kw["headers"] = self.headers

        if self._proxy_requires_url_absolute_form(u):
            response = conn.urlopen(method, url, **kw)
        else:
            response = conn.urlopen(method, u.request_uri, **kw)

        redirect_location = redirect and response.get_redirect_location()
        if not redirect_location:
            return response

        # Support relative URLs for redirecting.
        redirect_location = urljoin(url, redirect_location)

        if response.status == 303:
            # Change the method according to RFC 9110, Section 15.4.4.
            method = "GET"
            # And lose the body not to transfer anything sensitive.
            kw["body"] = None
            kw["headers"] = HTTPHeaderDict(kw["headers"])._prepare_for_method_change()

        retries = kw.get("retries")
        if not isinstance(retries, Retry):
            retries = Retry.from_int(retries, redirect=redirect)

        # Strip headers marked as unsafe to forward to the redirected location.
        # Check remove_headers_on_redirect to avoid a potential network call within
        # conn.is_same_host() which may use socket.gethostbyname() in the future.
        if retries.remove_headers_on_redirect and not conn.is_same_host(
            redirect_location
        ):
            new_headers = kw["headers"].copy()
            for header in kw["headers"]:
                if header.lower() in retries.remove_headers_on_redirect:
                    new_headers.pop(header, None)
            kw["headers"] = new_headers

        try:
            retries = retries.increment(method, url, response=response, _pool=conn)
        except MaxRetryError:
            if retries.raise_on_redirect:
                response.drain_conn()
                raise
            return response

        kw["retries"] = retries
        kw["redirect"] = redirect

        log.info("Redirecting %s -> %s", url, redirect_location)

        response.drain_conn()
        return self.urlopen(method, redirect_location, **kw)


class ProxyManager(PoolManager):
    """
    Behaves just like :class:`PoolManager`, but sends all requests through
    the defined proxy, using the CONNECT method for HTTPS URLs.

    :param proxy_url:
        The URL of the proxy to be used.

    :param proxy_headers:
        A dictionary containing headers that will be sent to the proxy. In case
        of HTTP they are being sent with each request, while in the
        HTTPS/CONNECT case they are sent only once. Could be used for proxy
        authentication.

    :param proxy_ssl_context:
        The proxy SSL context is used to establish the TLS connection to the
        proxy when using HTTPS proxies.

    :param use_forwarding_for_https:
        (Defaults to False) If set to True will forward requests to the HTTPS
        proxy to be made on behalf of the client instead of creating a TLS
        tunnel via the CONNECT method. **Enabling this flag means that request
        and response headers and content will be visible from the HTTPS proxy**
        whereas tunneling keeps request and response headers and content
        private.  IP address, target hostname, SNI, and port are always visible
        to an HTTPS proxy even when this flag is disabled.

    :param proxy_assert_hostname:
        The hostname of the certificate to verify against.

    :param proxy_assert_fingerprint:
        The fingerprint of the certificate to verify against.

    Example:

    .. code-block:: python

        import urllib3

        proxy = urllib3.ProxyManager("https://localhost:3128/")

        resp1 = proxy.request("GET", "https://google.com/")
        resp2 = proxy.request("GET", "https://httpbin.org/")

        print(len(proxy.pools))
        # 1

        resp3 = proxy.request("GET", "https://httpbin.org/")
        resp4 = proxy.request("GET", "https://twitter.com/")

        print(len(proxy.pools))
        # 3

    """

    def __init__(
        self,
        proxy_url: str,
        num_pools: int = 10,
        headers: typing.Mapping[str, str] | None = None,
        proxy_headers: typing.Mapping[str, str] | None = None,
        proxy_ssl_context: ssl.SSLContext | None = None,
        use_forwarding_for_https: bool = False,
        proxy_assert_hostname: None | str | typing.Literal[False] = None,
        proxy_assert_fingerprint: str | None = None,
        **connection_pool_kw: typing.Any,
    ) -> None:
        if isinstance(proxy_url, HTTPConnectionPool):
            str_proxy_url = f"{proxy_url.scheme}://{proxy_url.host}:{proxy_url.port}"
        else:
            str_proxy_url = proxy_url
        proxy = parse_url(str_proxy_url)

        if proxy.scheme not in ("http", "https"):
            raise ProxySchemeUnknown(proxy.scheme)

        if not proxy.port:
            port = port_by_scheme.get(proxy.scheme, 80)
            proxy = proxy._replace(port=port)

        self.proxy = proxy
        self.proxy_headers = proxy_headers or {}
        self.proxy_ssl_context = proxy_ssl_context
        self.proxy_config = ProxyConfig(
            proxy_ssl_context,
            use_forwarding_for_https,
            proxy_assert_hostname,
            proxy_assert_fingerprint,
        )

        connection_pool_kw["_proxy"] = self.proxy
        connection_pool_kw["_proxy_headers"] = self.proxy_headers
        connection_pool_kw["_proxy_config"] = self.proxy_config

        super().__init__(num_pools, headers, **connection_pool_kw)

    def connection_from_host(
        self,
        host: str | None,
        port: int | None = None,
        scheme: str | None = "http",
        pool_kwargs: dict[str, typing.Any] | None = None,
    ) -> HTTPConnectionPool:
        if scheme == "https":
            return super().connection_from_host(
                host, port, scheme, pool_kwargs=pool_kwargs
            )

        return super().connection_from_host(
            self.proxy.host, self.proxy.port, self.proxy.scheme, pool_kwargs=pool_kwargs  # type: ignore[union-attr]
        )

    def _set_proxy_headers(
        self, url: str, headers: typing.Mapping[str, str] | None = None
    ) -> typing.Mapping[str, str]:
        """
        Sets headers needed by proxies: specifically, the Accept and Host
        headers. Only sets headers not provided by the user.
        """
        headers_ = {"Accept": "*/*"}

        netloc = parse_url(url).netloc
        if netloc:
            headers_["Host"] = netloc

        if headers:
            headers_.update(headers)
        return headers_

    def urlopen(  # type: ignore[override]
        self, method: str, url: str, redirect: bool = True, **kw: typing.Any
    ) -> BaseHTTPResponse:
        "Same as HTTP(S)ConnectionPool.urlopen, ``url`` must be absolute."
        u = parse_url(url)
        if not connection_requires_http_tunnel(self.proxy, self.proxy_config, u.scheme):
            # For connections using HTTP CONNECT, httplib sets the necessary
            # headers on the CONNECT to the proxy. If we're not using CONNECT,
            # we'll definitely need to set 'Host' at the very least.
            headers = kw.get("headers", self.headers)
            kw["headers"] = self._set_proxy_headers(url, headers)

        return super().urlopen(method, url, redirect=redirect, **kw)


def proxy_from_url(url: str, **kw: typing.Any) -> ProxyManager:
    return ProxyManager(proxy_url=url, **kw)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\req\req_uninstall.py ===
import functools
import os
import sys
import sysconfig
from importlib.util import cache_from_source
from typing import Any, Callable, Dict, Generator, Iterable, List, Optional, Set, Tuple

from pip._internal.exceptions import LegacyDistutilsInstall, UninstallMissingRecord
from pip._internal.locations import get_bin_prefix, get_bin_user
from pip._internal.metadata import BaseDistribution
from pip._internal.utils.compat import WINDOWS
from pip._internal.utils.egg_link import egg_link_path_from_location
from pip._internal.utils.logging import getLogger, indent_log
from pip._internal.utils.misc import ask, normalize_path, renames, rmtree
from pip._internal.utils.temp_dir import AdjacentTempDirectory, TempDirectory
from pip._internal.utils.virtualenv import running_under_virtualenv

logger = getLogger(__name__)


def _script_names(
    bin_dir: str, script_name: str, is_gui: bool
) -> Generator[str, None, None]:
    """Create the fully qualified name of the files created by
    {console,gui}_scripts for the given ``dist``.
    Returns the list of file names
    """
    exe_name = os.path.join(bin_dir, script_name)
    yield exe_name
    if not WINDOWS:
        return
    yield f"{exe_name}.exe"
    yield f"{exe_name}.exe.manifest"
    if is_gui:
        yield f"{exe_name}-script.pyw"
    else:
        yield f"{exe_name}-script.py"


def _unique(
    fn: Callable[..., Generator[Any, None, None]],
) -> Callable[..., Generator[Any, None, None]]:
    @functools.wraps(fn)
    def unique(*args: Any, **kw: Any) -> Generator[Any, None, None]:
        seen: Set[Any] = set()
        for item in fn(*args, **kw):
            if item not in seen:
                seen.add(item)
                yield item

    return unique


@_unique
def uninstallation_paths(dist: BaseDistribution) -> Generator[str, None, None]:
    """
    Yield all the uninstallation paths for dist based on RECORD-without-.py[co]

    Yield paths to all the files in RECORD. For each .py file in RECORD, add
    the .pyc and .pyo in the same directory.

    UninstallPathSet.add() takes care of the __pycache__ .py[co].

    If RECORD is not found, raises an error,
    with possible information from the INSTALLER file.

    https://packaging.python.org/specifications/recording-installed-packages/
    """
    location = dist.location
    assert location is not None, "not installed"

    entries = dist.iter_declared_entries()
    if entries is None:
        raise UninstallMissingRecord(distribution=dist)

    for entry in entries:
        path = os.path.join(location, entry)
        yield path
        if path.endswith(".py"):
            dn, fn = os.path.split(path)
            base = fn[:-3]
            path = os.path.join(dn, base + ".pyc")
            yield path
            path = os.path.join(dn, base + ".pyo")
            yield path


def compact(paths: Iterable[str]) -> Set[str]:
    """Compact a path set to contain the minimal number of paths
    necessary to contain all paths in the set. If /a/path/ and
    /a/path/to/a/file.txt are both in the set, leave only the
    shorter path."""

    sep = os.path.sep
    short_paths: Set[str] = set()
    for path in sorted(paths, key=len):
        should_skip = any(
            path.startswith(shortpath.rstrip("*"))
            and path[len(shortpath.rstrip("*").rstrip(sep))] == sep
            for shortpath in short_paths
        )
        if not should_skip:
            short_paths.add(path)
    return short_paths


def compress_for_rename(paths: Iterable[str]) -> Set[str]:
    """Returns a set containing the paths that need to be renamed.

    This set may include directories when the original sequence of paths
    included every file on disk.
    """
    case_map = {os.path.normcase(p): p for p in paths}
    remaining = set(case_map)
    unchecked = sorted({os.path.split(p)[0] for p in case_map.values()}, key=len)
    wildcards: Set[str] = set()

    def norm_join(*a: str) -> str:
        return os.path.normcase(os.path.join(*a))

    for root in unchecked:
        if any(os.path.normcase(root).startswith(w) for w in wildcards):
            # This directory has already been handled.
            continue

        all_files: Set[str] = set()
        all_subdirs: Set[str] = set()
        for dirname, subdirs, files in os.walk(root):
            all_subdirs.update(norm_join(root, dirname, d) for d in subdirs)
            all_files.update(norm_join(root, dirname, f) for f in files)
        # If all the files we found are in our remaining set of files to
        # remove, then remove them from the latter set and add a wildcard
        # for the directory.
        if not (all_files - remaining):
            remaining.difference_update(all_files)
            wildcards.add(root + os.sep)

    return set(map(case_map.__getitem__, remaining)) | wildcards


def compress_for_output_listing(paths: Iterable[str]) -> Tuple[Set[str], Set[str]]:
    """Returns a tuple of 2 sets of which paths to display to user

    The first set contains paths that would be deleted. Files of a package
    are not added and the top-level directory of the package has a '*' added
    at the end - to signify that all it's contents are removed.

    The second set contains files that would have been skipped in the above
    folders.
    """

    will_remove = set(paths)
    will_skip = set()

    # Determine folders and files
    folders = set()
    files = set()
    for path in will_remove:
        if path.endswith(".pyc"):
            continue
        if path.endswith("__init__.py") or ".dist-info" in path:
            folders.add(os.path.dirname(path))
        files.add(path)

    _normcased_files = set(map(os.path.normcase, files))

    folders = compact(folders)

    # This walks the tree using os.walk to not miss extra folders
    # that might get added.
    for folder in folders:
        for dirpath, _, dirfiles in os.walk(folder):
            for fname in dirfiles:
                if fname.endswith(".pyc"):
                    continue

                file_ = os.path.join(dirpath, fname)
                if (
                    os.path.isfile(file_)
                    and os.path.normcase(file_) not in _normcased_files
                ):
                    # We are skipping this file. Add it to the set.
                    will_skip.add(file_)

    will_remove = files | {os.path.join(folder, "*") for folder in folders}

    return will_remove, will_skip


class StashedUninstallPathSet:
    """A set of file rename operations to stash files while
    tentatively uninstalling them."""

    def __init__(self) -> None:
        # Mapping from source file root to [Adjacent]TempDirectory
        # for files under that directory.
        self._save_dirs: Dict[str, TempDirectory] = {}
        # (old path, new path) tuples for each move that may need
        # to be undone.
        self._moves: List[Tuple[str, str]] = []

    def _get_directory_stash(self, path: str) -> str:
        """Stashes a directory.

        Directories are stashed adjacent to their original location if
        possible, or else moved/copied into the user's temp dir."""

        try:
            save_dir: TempDirectory = AdjacentTempDirectory(path)
        except OSError:
            save_dir = TempDirectory(kind="uninstall")
        self._save_dirs[os.path.normcase(path)] = save_dir

        return save_dir.path

    def _get_file_stash(self, path: str) -> str:
        """Stashes a file.

        If no root has been provided, one will be created for the directory
        in the user's temp directory."""
        path = os.path.normcase(path)
        head, old_head = os.path.dirname(path), None
        save_dir = None

        while head != old_head:
            try:
                save_dir = self._save_dirs[head]
                break
            except KeyError:
                pass
            head, old_head = os.path.dirname(head), head
        else:
            # Did not find any suitable root
            head = os.path.dirname(path)
            save_dir = TempDirectory(kind="uninstall")
            self._save_dirs[head] = save_dir

        relpath = os.path.relpath(path, head)
        if relpath and relpath != os.path.curdir:
            return os.path.join(save_dir.path, relpath)
        return save_dir.path

    def stash(self, path: str) -> str:
        """Stashes the directory or file and returns its new location.
        Handle symlinks as files to avoid modifying the symlink targets.
        """
        path_is_dir = os.path.isdir(path) and not os.path.islink(path)
        if path_is_dir:
            new_path = self._get_directory_stash(path)
        else:
            new_path = self._get_file_stash(path)

        self._moves.append((path, new_path))
        if path_is_dir and os.path.isdir(new_path):
            # If we're moving a directory, we need to
            # remove the destination first or else it will be
            # moved to inside the existing directory.
            # We just created new_path ourselves, so it will
            # be removable.
            os.rmdir(new_path)
        renames(path, new_path)
        return new_path

    def commit(self) -> None:
        """Commits the uninstall by removing stashed files."""
        for save_dir in self._save_dirs.values():
            save_dir.cleanup()
        self._moves = []
        self._save_dirs = {}

    def rollback(self) -> None:
        """Undoes the uninstall by moving stashed files back."""
        for p in self._moves:
            logger.info("Moving to %s\n from %s", *p)

        for new_path, path in self._moves:
            try:
                logger.debug("Replacing %s from %s", new_path, path)
                if os.path.isfile(new_path) or os.path.islink(new_path):
                    os.unlink(new_path)
                elif os.path.isdir(new_path):
                    rmtree(new_path)
                renames(path, new_path)
            except OSError as ex:
                logger.error("Failed to restore %s", new_path)
                logger.debug("Exception: %s", ex)

        self.commit()

    @property
    def can_rollback(self) -> bool:
        return bool(self._moves)


class UninstallPathSet:
    """A set of file paths to be removed in the uninstallation of a
    requirement."""

    def __init__(self, dist: BaseDistribution) -> None:
        self._paths: Set[str] = set()
        self._refuse: Set[str] = set()
        self._pth: Dict[str, UninstallPthEntries] = {}
        self._dist = dist
        self._moved_paths = StashedUninstallPathSet()
        # Create local cache of normalize_path results. Creating an UninstallPathSet
        # can result in hundreds/thousands of redundant calls to normalize_path with
        # the same args, which hurts performance.
        self._normalize_path_cached = functools.lru_cache(normalize_path)

    def _permitted(self, path: str) -> bool:
        """
        Return True if the given path is one we are permitted to
        remove/modify, False otherwise.

        """
        # aka is_local, but caching normalized sys.prefix
        if not running_under_virtualenv():
            return True
        return path.startswith(self._normalize_path_cached(sys.prefix))

    def add(self, path: str) -> None:
        head, tail = os.path.split(path)

        # we normalize the head to resolve parent directory symlinks, but not
        # the tail, since we only want to uninstall symlinks, not their targets
        path = os.path.join(self._normalize_path_cached(head), os.path.normcase(tail))

        if not os.path.exists(path):
            return
        if self._permitted(path):
            self._paths.add(path)
        else:
            self._refuse.add(path)

        # __pycache__ files can show up after 'installed-files.txt' is created,
        # due to imports
        if os.path.splitext(path)[1] == ".py":
            self.add(cache_from_source(path))

    def add_pth(self, pth_file: str, entry: str) -> None:
        pth_file = self._normalize_path_cached(pth_file)
        if self._permitted(pth_file):
            if pth_file not in self._pth:
                self._pth[pth_file] = UninstallPthEntries(pth_file)
            self._pth[pth_file].add(entry)
        else:
            self._refuse.add(pth_file)

    def remove(self, auto_confirm: bool = False, verbose: bool = False) -> None:
        """Remove paths in ``self._paths`` with confirmation (unless
        ``auto_confirm`` is True)."""

        if not self._paths:
            logger.info(
                "Can't uninstall '%s'. No files were found to uninstall.",
                self._dist.raw_name,
            )
            return

        dist_name_version = f"{self._dist.raw_name}-{self._dist.raw_version}"
        logger.info("Uninstalling %s:", dist_name_version)

        with indent_log():
            if auto_confirm or self._allowed_to_proceed(verbose):
                moved = self._moved_paths

                for_rename = compress_for_rename(self._paths)

                for path in sorted(compact(for_rename)):
                    moved.stash(path)
                    logger.verbose("Removing file or directory %s", path)

                for pth in self._pth.values():
                    pth.remove()

                logger.info("Successfully uninstalled %s", dist_name_version)

    def _allowed_to_proceed(self, verbose: bool) -> bool:
        """Display which files would be deleted and prompt for confirmation"""

        def _display(msg: str, paths: Iterable[str]) -> None:
            if not paths:
                return

            logger.info(msg)
            with indent_log():
                for path in sorted(compact(paths)):
                    logger.info(path)

        if not verbose:
            will_remove, will_skip = compress_for_output_listing(self._paths)
        else:
            # In verbose mode, display all the files that are going to be
            # deleted.
            will_remove = set(self._paths)
            will_skip = set()

        _display("Would remove:", will_remove)
        _display("Would not remove (might be manually added):", will_skip)
        _display("Would not remove (outside of prefix):", self._refuse)
        if verbose:
            _display("Will actually move:", compress_for_rename(self._paths))

        return ask("Proceed (Y/n)? ", ("y", "n", "")) != "n"

    def rollback(self) -> None:
        """Rollback the changes previously made by remove()."""
        if not self._moved_paths.can_rollback:
            logger.error(
                "Can't roll back %s; was not uninstalled",
                self._dist.raw_name,
            )
            return
        logger.info("Rolling back uninstall of %s", self._dist.raw_name)
        self._moved_paths.rollback()
        for pth in self._pth.values():
            pth.rollback()

    def commit(self) -> None:
        """Remove temporary save dir: rollback will no longer be possible."""
        self._moved_paths.commit()

    @classmethod
    def from_dist(cls, dist: BaseDistribution) -> "UninstallPathSet":
        dist_location = dist.location
        info_location = dist.info_location
        if dist_location is None:
            logger.info(
                "Not uninstalling %s since it is not installed",
                dist.canonical_name,
            )
            return cls(dist)

        normalized_dist_location = normalize_path(dist_location)
        if not dist.local:
            logger.info(
                "Not uninstalling %s at %s, outside environment %s",
                dist.canonical_name,
                normalized_dist_location,
                sys.prefix,
            )
            return cls(dist)

        if normalized_dist_location in {
            p
            for p in {sysconfig.get_path("stdlib"), sysconfig.get_path("platstdlib")}
            if p
        }:
            logger.info(
                "Not uninstalling %s at %s, as it is in the standard library.",
                dist.canonical_name,
                normalized_dist_location,
            )
            return cls(dist)

        paths_to_remove = cls(dist)
        develop_egg_link = egg_link_path_from_location(dist.raw_name)

        # Distribution is installed with metadata in a "flat" .egg-info
        # directory. This means it is not a modern .dist-info installation, an
        # egg, or legacy editable.
        setuptools_flat_installation = (
            dist.installed_with_setuptools_egg_info
            and info_location is not None
            and os.path.exists(info_location)
            # If dist is editable and the location points to a ``.egg-info``,
            # we are in fact in the legacy editable case.
            and not info_location.endswith(f"{dist.setuptools_filename}.egg-info")
        )

        # Uninstall cases order do matter as in the case of 2 installs of the
        # same package, pip needs to uninstall the currently detected version
        if setuptools_flat_installation:
            if info_location is not None:
                paths_to_remove.add(info_location)
            installed_files = dist.iter_declared_entries()
            if installed_files is not None:
                for installed_file in installed_files:
                    paths_to_remove.add(os.path.join(dist_location, installed_file))
            # FIXME: need a test for this elif block
            # occurs with --single-version-externally-managed/--record outside
            # of pip
            elif dist.is_file("top_level.txt"):
                try:
                    namespace_packages = dist.read_text("namespace_packages.txt")
                except FileNotFoundError:
                    namespaces = []
                else:
                    namespaces = namespace_packages.splitlines(keepends=False)
                for top_level_pkg in [
                    p
                    for p in dist.read_text("top_level.txt").splitlines()
                    if p and p not in namespaces
                ]:
                    path = os.path.join(dist_location, top_level_pkg)
                    paths_to_remove.add(path)
                    paths_to_remove.add(f"{path}.py")
                    paths_to_remove.add(f"{path}.pyc")
                    paths_to_remove.add(f"{path}.pyo")

        elif dist.installed_by_distutils:
            raise LegacyDistutilsInstall(distribution=dist)

        elif dist.installed_as_egg:
            # package installed by easy_install
            # We cannot match on dist.egg_name because it can slightly vary
            # i.e. setuptools-0.6c11-py2.6.egg vs setuptools-0.6rc11-py2.6.egg
            # XXX We use normalized_dist_location because dist_location my contain
            # a trailing / if the distribution is a zipped egg
            # (which is not a directory).
            paths_to_remove.add(normalized_dist_location)
            easy_install_egg = os.path.split(normalized_dist_location)[1]
            easy_install_pth = os.path.join(
                os.path.dirname(normalized_dist_location),
                "easy-install.pth",
            )
            paths_to_remove.add_pth(easy_install_pth, "./" + easy_install_egg)

        elif dist.installed_with_dist_info:
            for path in uninstallation_paths(dist):
                paths_to_remove.add(path)

        elif develop_egg_link:
            # PEP 660 modern editable is handled in the ``.dist-info`` case
            # above, so this only covers the setuptools-style editable.
            with open(develop_egg_link) as fh:
                link_pointer = os.path.normcase(fh.readline().strip())
                normalized_link_pointer = paths_to_remove._normalize_path_cached(
                    link_pointer
                )
            assert os.path.samefile(
                normalized_link_pointer, normalized_dist_location
            ), (
                f"Egg-link {develop_egg_link} (to {link_pointer}) does not match "
                f"installed location of {dist.raw_name} (at {dist_location})"
            )
            paths_to_remove.add(develop_egg_link)
            easy_install_pth = os.path.join(
                os.path.dirname(develop_egg_link), "easy-install.pth"
            )
            paths_to_remove.add_pth(easy_install_pth, dist_location)

        else:
            logger.debug(
                "Not sure how to uninstall: %s - Check: %s",
                dist,
                dist_location,
            )

        if dist.in_usersite:
            bin_dir = get_bin_user()
        else:
            bin_dir = get_bin_prefix()

        # find distutils scripts= scripts
        try:
            for script in dist.iter_distutils_script_names():
                paths_to_remove.add(os.path.join(bin_dir, script))
                if WINDOWS:
                    paths_to_remove.add(os.path.join(bin_dir, f"{script}.bat"))
        except (FileNotFoundError, NotADirectoryError):
            pass

        # find console_scripts and gui_scripts
        def iter_scripts_to_remove(
            dist: BaseDistribution,
            bin_dir: str,
        ) -> Generator[str, None, None]:
            for entry_point in dist.iter_entry_points():
                if entry_point.group == "console_scripts":
                    yield from _script_names(bin_dir, entry_point.name, False)
                elif entry_point.group == "gui_scripts":
                    yield from _script_names(bin_dir, entry_point.name, True)

        for s in iter_scripts_to_remove(dist, bin_dir):
            paths_to_remove.add(s)

        return paths_to_remove


class UninstallPthEntries:
    def __init__(self, pth_file: str) -> None:
        self.file = pth_file
        self.entries: Set[str] = set()
        self._saved_lines: Optional[List[bytes]] = None

    def add(self, entry: str) -> None:
        entry = os.path.normcase(entry)
        # On Windows, os.path.normcase converts the entry to use
        # backslashes.  This is correct for entries that describe absolute
        # paths outside of site-packages, but all the others use forward
        # slashes.
        # os.path.splitdrive is used instead of os.path.isabs because isabs
        # treats non-absolute paths with drive letter markings like c:foo\bar
        # as absolute paths. It also does not recognize UNC paths if they don't
        # have more than "\\sever\share". Valid examples: "\\server\share\" or
        # "\\server\share\folder".
        if WINDOWS and not os.path.splitdrive(entry)[0]:
            entry = entry.replace("\\", "/")
        self.entries.add(entry)

    def remove(self) -> None:
        logger.verbose("Removing pth entries from %s:", self.file)

        # If the file doesn't exist, log a warning and return
        if not os.path.isfile(self.file):
            logger.warning("Cannot remove entries from nonexistent file %s", self.file)
            return
        with open(self.file, "rb") as fh:
            # windows uses '\r\n' with py3k, but uses '\n' with py2.x
            lines = fh.readlines()
            self._saved_lines = lines
        if any(b"\r\n" in line for line in lines):
            endline = "\r\n"
        else:
            endline = "\n"
        # handle missing trailing newline
        if lines and not lines[-1].endswith(endline.encode("utf-8")):
            lines[-1] = lines[-1] + endline.encode("utf-8")
        for entry in self.entries:
            try:
                logger.verbose("Removing entry: %s", entry)
                lines.remove((entry + endline).encode("utf-8"))
            except ValueError:
                pass
        with open(self.file, "wb") as fh:
            fh.writelines(lines)

    def rollback(self) -> bool:
        if self._saved_lines is None:
            logger.error("Cannot roll back changes to %s, none were made", self.file)
            return False
        logger.debug("Rolling %s back to previous state", self.file)
        with open(self.file, "wb") as fh:
            fh.writelines(self._saved_lines)
        return True

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\charset_normalizer\md.py ===
from __future__ import annotations

from functools import lru_cache
from logging import getLogger

from .constant import (
    COMMON_SAFE_ASCII_CHARACTERS,
    TRACE,
    UNICODE_SECONDARY_RANGE_KEYWORD,
)
from .utils import (
    is_accentuated,
    is_arabic,
    is_arabic_isolated_form,
    is_case_variable,
    is_cjk,
    is_emoticon,
    is_hangul,
    is_hiragana,
    is_katakana,
    is_latin,
    is_punctuation,
    is_separator,
    is_symbol,
    is_thai,
    is_unprintable,
    remove_accent,
    unicode_range,
    is_cjk_uncommon,
)


class MessDetectorPlugin:
    """
    Base abstract class used for mess detection plugins.
    All detectors MUST extend and implement given methods.
    """

    def eligible(self, character: str) -> bool:
        """
        Determine if given character should be fed in.
        """
        raise NotImplementedError  # pragma: nocover

    def feed(self, character: str) -> None:
        """
        The main routine to be executed upon character.
        Insert the logic in witch the text would be considered chaotic.
        """
        raise NotImplementedError  # pragma: nocover

    def reset(self) -> None:  # pragma: no cover
        """
        Permit to reset the plugin to the initial state.
        """
        raise NotImplementedError

    @property
    def ratio(self) -> float:
        """
        Compute the chaos ratio based on what your feed() has seen.
        Must NOT be lower than 0.; No restriction gt 0.
        """
        raise NotImplementedError  # pragma: nocover


class TooManySymbolOrPunctuationPlugin(MessDetectorPlugin):
    def __init__(self) -> None:
        self._punctuation_count: int = 0
        self._symbol_count: int = 0
        self._character_count: int = 0

        self._last_printable_char: str | None = None
        self._frenzy_symbol_in_word: bool = False

    def eligible(self, character: str) -> bool:
        return character.isprintable()

    def feed(self, character: str) -> None:
        self._character_count += 1

        if (
            character != self._last_printable_char
            and character not in COMMON_SAFE_ASCII_CHARACTERS
        ):
            if is_punctuation(character):
                self._punctuation_count += 1
            elif (
                character.isdigit() is False
                and is_symbol(character)
                and is_emoticon(character) is False
            ):
                self._symbol_count += 2

        self._last_printable_char = character

    def reset(self) -> None:  # Abstract
        self._punctuation_count = 0
        self._character_count = 0
        self._symbol_count = 0

    @property
    def ratio(self) -> float:
        if self._character_count == 0:
            return 0.0

        ratio_of_punctuation: float = (
            self._punctuation_count + self._symbol_count
        ) / self._character_count

        return ratio_of_punctuation if ratio_of_punctuation >= 0.3 else 0.0


class TooManyAccentuatedPlugin(MessDetectorPlugin):
    def __init__(self) -> None:
        self._character_count: int = 0
        self._accentuated_count: int = 0

    def eligible(self, character: str) -> bool:
        return character.isalpha()

    def feed(self, character: str) -> None:
        self._character_count += 1

        if is_accentuated(character):
            self._accentuated_count += 1

    def reset(self) -> None:  # Abstract
        self._character_count = 0
        self._accentuated_count = 0

    @property
    def ratio(self) -> float:
        if self._character_count < 8:
            return 0.0

        ratio_of_accentuation: float = self._accentuated_count / self._character_count
        return ratio_of_accentuation if ratio_of_accentuation >= 0.35 else 0.0


class UnprintablePlugin(MessDetectorPlugin):
    def __init__(self) -> None:
        self._unprintable_count: int = 0
        self._character_count: int = 0

    def eligible(self, character: str) -> bool:
        return True

    def feed(self, character: str) -> None:
        if is_unprintable(character):
            self._unprintable_count += 1
        self._character_count += 1

    def reset(self) -> None:  # Abstract
        self._unprintable_count = 0

    @property
    def ratio(self) -> float:
        if self._character_count == 0:
            return 0.0

        return (self._unprintable_count * 8) / self._character_count


class SuspiciousDuplicateAccentPlugin(MessDetectorPlugin):
    def __init__(self) -> None:
        self._successive_count: int = 0
        self._character_count: int = 0

        self._last_latin_character: str | None = None

    def eligible(self, character: str) -> bool:
        return character.isalpha() and is_latin(character)

    def feed(self, character: str) -> None:
        self._character_count += 1
        if (
            self._last_latin_character is not None
            and is_accentuated(character)
            and is_accentuated(self._last_latin_character)
        ):
            if character.isupper() and self._last_latin_character.isupper():
                self._successive_count += 1
            # Worse if its the same char duplicated with different accent.
            if remove_accent(character) == remove_accent(self._last_latin_character):
                self._successive_count += 1
        self._last_latin_character = character

    def reset(self) -> None:  # Abstract
        self._successive_count = 0
        self._character_count = 0
        self._last_latin_character = None

    @property
    def ratio(self) -> float:
        if self._character_count == 0:
            return 0.0

        return (self._successive_count * 2) / self._character_count


class SuspiciousRange(MessDetectorPlugin):
    def __init__(self) -> None:
        self._suspicious_successive_range_count: int = 0
        self._character_count: int = 0
        self._last_printable_seen: str | None = None

    def eligible(self, character: str) -> bool:
        return character.isprintable()

    def feed(self, character: str) -> None:
        self._character_count += 1

        if (
            character.isspace()
            or is_punctuation(character)
            or character in COMMON_SAFE_ASCII_CHARACTERS
        ):
            self._last_printable_seen = None
            return

        if self._last_printable_seen is None:
            self._last_printable_seen = character
            return

        unicode_range_a: str | None = unicode_range(self._last_printable_seen)
        unicode_range_b: str | None = unicode_range(character)

        if is_suspiciously_successive_range(unicode_range_a, unicode_range_b):
            self._suspicious_successive_range_count += 1

        self._last_printable_seen = character

    def reset(self) -> None:  # Abstract
        self._character_count = 0
        self._suspicious_successive_range_count = 0
        self._last_printable_seen = None

    @property
    def ratio(self) -> float:
        if self._character_count <= 13:
            return 0.0

        ratio_of_suspicious_range_usage: float = (
            self._suspicious_successive_range_count * 2
        ) / self._character_count

        return ratio_of_suspicious_range_usage


class SuperWeirdWordPlugin(MessDetectorPlugin):
    def __init__(self) -> None:
        self._word_count: int = 0
        self._bad_word_count: int = 0
        self._foreign_long_count: int = 0

        self._is_current_word_bad: bool = False
        self._foreign_long_watch: bool = False

        self._character_count: int = 0
        self._bad_character_count: int = 0

        self._buffer: str = ""
        self._buffer_accent_count: int = 0
        self._buffer_glyph_count: int = 0

    def eligible(self, character: str) -> bool:
        return True

    def feed(self, character: str) -> None:
        if character.isalpha():
            self._buffer += character
            if is_accentuated(character):
                self._buffer_accent_count += 1
            if (
                self._foreign_long_watch is False
                and (is_latin(character) is False or is_accentuated(character))
                and is_cjk(character) is False
                and is_hangul(character) is False
                and is_katakana(character) is False
                and is_hiragana(character) is False
                and is_thai(character) is False
            ):
                self._foreign_long_watch = True
            if (
                is_cjk(character)
                or is_hangul(character)
                or is_katakana(character)
                or is_hiragana(character)
                or is_thai(character)
            ):
                self._buffer_glyph_count += 1
            return
        if not self._buffer:
            return
        if (
            character.isspace() or is_punctuation(character) or is_separator(character)
        ) and self._buffer:
            self._word_count += 1
            buffer_length: int = len(self._buffer)

            self._character_count += buffer_length

            if buffer_length >= 4:
                if self._buffer_accent_count / buffer_length >= 0.5:
                    self._is_current_word_bad = True
                # Word/Buffer ending with an upper case accentuated letter are so rare,
                # that we will consider them all as suspicious. Same weight as foreign_long suspicious.
                elif (
                    is_accentuated(self._buffer[-1])
                    and self._buffer[-1].isupper()
                    and all(_.isupper() for _ in self._buffer) is False
                ):
                    self._foreign_long_count += 1
                    self._is_current_word_bad = True
                elif self._buffer_glyph_count == 1:
                    self._is_current_word_bad = True
                    self._foreign_long_count += 1
            if buffer_length >= 24 and self._foreign_long_watch:
                camel_case_dst = [
                    i
                    for c, i in zip(self._buffer, range(0, buffer_length))
                    if c.isupper()
                ]
                probable_camel_cased: bool = False

                if camel_case_dst and (len(camel_case_dst) / buffer_length <= 0.3):
                    probable_camel_cased = True

                if not probable_camel_cased:
                    self._foreign_long_count += 1
                    self._is_current_word_bad = True

            if self._is_current_word_bad:
                self._bad_word_count += 1
                self._bad_character_count += len(self._buffer)
                self._is_current_word_bad = False

            self._foreign_long_watch = False
            self._buffer = ""
            self._buffer_accent_count = 0
            self._buffer_glyph_count = 0
        elif (
            character not in {"<", ">", "-", "=", "~", "|", "_"}
            and character.isdigit() is False
            and is_symbol(character)
        ):
            self._is_current_word_bad = True
            self._buffer += character

    def reset(self) -> None:  # Abstract
        self._buffer = ""
        self._is_current_word_bad = False
        self._foreign_long_watch = False
        self._bad_word_count = 0
        self._word_count = 0
        self._character_count = 0
        self._bad_character_count = 0
        self._foreign_long_count = 0

    @property
    def ratio(self) -> float:
        if self._word_count <= 10 and self._foreign_long_count == 0:
            return 0.0

        return self._bad_character_count / self._character_count


class CjkUncommonPlugin(MessDetectorPlugin):
    """
    Detect messy CJK text that probably means nothing.
    """

    def __init__(self) -> None:
        self._character_count: int = 0
        self._uncommon_count: int = 0

    def eligible(self, character: str) -> bool:
        return is_cjk(character)

    def feed(self, character: str) -> None:
        self._character_count += 1

        if is_cjk_uncommon(character):
            self._uncommon_count += 1
            return

    def reset(self) -> None:  # Abstract
        self._character_count = 0
        self._uncommon_count = 0

    @property
    def ratio(self) -> float:
        if self._character_count < 8:
            return 0.0

        uncommon_form_usage: float = self._uncommon_count / self._character_count

        # we can be pretty sure it's garbage when uncommon characters are widely
        # used. otherwise it could just be traditional chinese for example.
        return uncommon_form_usage / 10 if uncommon_form_usage > 0.5 else 0.0


class ArchaicUpperLowerPlugin(MessDetectorPlugin):
    def __init__(self) -> None:
        self._buf: bool = False

        self._character_count_since_last_sep: int = 0

        self._successive_upper_lower_count: int = 0
        self._successive_upper_lower_count_final: int = 0

        self._character_count: int = 0

        self._last_alpha_seen: str | None = None
        self._current_ascii_only: bool = True

    def eligible(self, character: str) -> bool:
        return True

    def feed(self, character: str) -> None:
        is_concerned = character.isalpha() and is_case_variable(character)
        chunk_sep = is_concerned is False

        if chunk_sep and self._character_count_since_last_sep > 0:
            if (
                self._character_count_since_last_sep <= 64
                and character.isdigit() is False
                and self._current_ascii_only is False
            ):
                self._successive_upper_lower_count_final += (
                    self._successive_upper_lower_count
                )

            self._successive_upper_lower_count = 0
            self._character_count_since_last_sep = 0
            self._last_alpha_seen = None
            self._buf = False
            self._character_count += 1
            self._current_ascii_only = True

            return

        if self._current_ascii_only is True and character.isascii() is False:
            self._current_ascii_only = False

        if self._last_alpha_seen is not None:
            if (character.isupper() and self._last_alpha_seen.islower()) or (
                character.islower() and self._last_alpha_seen.isupper()
            ):
                if self._buf is True:
                    self._successive_upper_lower_count += 2
                    self._buf = False
                else:
                    self._buf = True
            else:
                self._buf = False

        self._character_count += 1
        self._character_count_since_last_sep += 1
        self._last_alpha_seen = character

    def reset(self) -> None:  # Abstract
        self._character_count = 0
        self._character_count_since_last_sep = 0
        self._successive_upper_lower_count = 0
        self._successive_upper_lower_count_final = 0
        self._last_alpha_seen = None
        self._buf = False
        self._current_ascii_only = True

    @property
    def ratio(self) -> float:
        if self._character_count == 0:
            return 0.0

        return self._successive_upper_lower_count_final / self._character_count


class ArabicIsolatedFormPlugin(MessDetectorPlugin):
    def __init__(self) -> None:
        self._character_count: int = 0
        self._isolated_form_count: int = 0

    def reset(self) -> None:  # Abstract
        self._character_count = 0
        self._isolated_form_count = 0

    def eligible(self, character: str) -> bool:
        return is_arabic(character)

    def feed(self, character: str) -> None:
        self._character_count += 1

        if is_arabic_isolated_form(character):
            self._isolated_form_count += 1

    @property
    def ratio(self) -> float:
        if self._character_count < 8:
            return 0.0

        isolated_form_usage: float = self._isolated_form_count / self._character_count

        return isolated_form_usage


@lru_cache(maxsize=1024)
def is_suspiciously_successive_range(
    unicode_range_a: str | None, unicode_range_b: str | None
) -> bool:
    """
    Determine if two Unicode range seen next to each other can be considered as suspicious.
    """
    if unicode_range_a is None or unicode_range_b is None:
        return True

    if unicode_range_a == unicode_range_b:
        return False

    if "Latin" in unicode_range_a and "Latin" in unicode_range_b:
        return False

    if "Emoticons" in unicode_range_a or "Emoticons" in unicode_range_b:
        return False

    # Latin characters can be accompanied with a combining diacritical mark
    # eg. Vietnamese.
    if ("Latin" in unicode_range_a or "Latin" in unicode_range_b) and (
        "Combining" in unicode_range_a or "Combining" in unicode_range_b
    ):
        return False

    keywords_range_a, keywords_range_b = (
        unicode_range_a.split(" "),
        unicode_range_b.split(" "),
    )

    for el in keywords_range_a:
        if el in UNICODE_SECONDARY_RANGE_KEYWORD:
            continue
        if el in keywords_range_b:
            return False

    # Japanese Exception
    range_a_jp_chars, range_b_jp_chars = (
        unicode_range_a
        in (
            "Hiragana",
            "Katakana",
        ),
        unicode_range_b in ("Hiragana", "Katakana"),
    )
    if (range_a_jp_chars or range_b_jp_chars) and (
        "CJK" in unicode_range_a or "CJK" in unicode_range_b
    ):
        return False
    if range_a_jp_chars and range_b_jp_chars:
        return False

    if "Hangul" in unicode_range_a or "Hangul" in unicode_range_b:
        if "CJK" in unicode_range_a or "CJK" in unicode_range_b:
            return False
        if unicode_range_a == "Basic Latin" or unicode_range_b == "Basic Latin":
            return False

    # Chinese/Japanese use dedicated range for punctuation and/or separators.
    if ("CJK" in unicode_range_a or "CJK" in unicode_range_b) or (
        unicode_range_a in ["Katakana", "Hiragana"]
        and unicode_range_b in ["Katakana", "Hiragana"]
    ):
        if "Punctuation" in unicode_range_a or "Punctuation" in unicode_range_b:
            return False
        if "Forms" in unicode_range_a or "Forms" in unicode_range_b:
            return False
        if unicode_range_a == "Basic Latin" or unicode_range_b == "Basic Latin":
            return False

    return True


@lru_cache(maxsize=2048)
def mess_ratio(
    decoded_sequence: str, maximum_threshold: float = 0.2, debug: bool = False
) -> float:
    """
    Compute a mess ratio given a decoded bytes sequence. The maximum threshold does stop the computation earlier.
    """

    detectors: list[MessDetectorPlugin] = [
        md_class() for md_class in MessDetectorPlugin.__subclasses__()
    ]

    length: int = len(decoded_sequence) + 1

    mean_mess_ratio: float = 0.0

    if length < 512:
        intermediary_mean_mess_ratio_calc: int = 32
    elif length <= 1024:
        intermediary_mean_mess_ratio_calc = 64
    else:
        intermediary_mean_mess_ratio_calc = 128

    for character, index in zip(decoded_sequence + "\n", range(length)):
        for detector in detectors:
            if detector.eligible(character):
                detector.feed(character)

        if (
            index > 0 and index % intermediary_mean_mess_ratio_calc == 0
        ) or index == length - 1:
            mean_mess_ratio = sum(dt.ratio for dt in detectors)

            if mean_mess_ratio >= maximum_threshold:
                break

    if debug:
        logger = getLogger("charset_normalizer")

        logger.log(
            TRACE,
            "Mess-detector extended-analysis start. "
            f"intermediary_mean_mess_ratio_calc={intermediary_mean_mess_ratio_calc} mean_mess_ratio={mean_mess_ratio} "
            f"maximum_threshold={maximum_threshold}",
        )

        if len(decoded_sequence) > 16:
            logger.log(TRACE, f"Starting with: {decoded_sequence[:16]}")
            logger.log(TRACE, f"Ending with: {decoded_sequence[-16::]}")

        for dt in detectors:
            logger.log(TRACE, f"{dt.__class__}: {dt.ratio}")

    return round(mean_mess_ratio, 3)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\_testing\__init__.py ===
from __future__ import annotations

from decimal import Decimal
import operator
import os
from sys import byteorder
from typing import (
    TYPE_CHECKING,
    Callable,
    ContextManager,
)
import warnings

import numpy as np

from pandas._config import using_string_dtype
from pandas._config.localization import (
    can_set_locale,
    get_locales,
    set_locale,
)

from pandas.compat import pa_version_under10p1

import pandas as pd
from pandas import (
    ArrowDtype,
    DataFrame,
    Index,
    MultiIndex,
    RangeIndex,
    Series,
)
from pandas._testing._io import (
    round_trip_localpath,
    round_trip_pathlib,
    round_trip_pickle,
    write_to_compressed,
)
from pandas._testing._warnings import (
    assert_produces_warning,
    maybe_produces_warning,
)
from pandas._testing.asserters import (
    assert_almost_equal,
    assert_attr_equal,
    assert_categorical_equal,
    assert_class_equal,
    assert_contains_all,
    assert_copy,
    assert_datetime_array_equal,
    assert_dict_equal,
    assert_equal,
    assert_extension_array_equal,
    assert_frame_equal,
    assert_index_equal,
    assert_indexing_slices_equivalent,
    assert_interval_array_equal,
    assert_is_sorted,
    assert_is_valid_plot_return_object,
    assert_metadata_equivalent,
    assert_numpy_array_equal,
    assert_period_array_equal,
    assert_series_equal,
    assert_sp_array_equal,
    assert_timedelta_array_equal,
    raise_assert_detail,
)
from pandas._testing.compat import (
    get_dtype,
    get_obj,
)
from pandas._testing.contexts import (
    assert_cow_warning,
    decompress_file,
    ensure_clean,
    raises_chained_assignment_error,
    set_timezone,
    use_numexpr,
    with_csv_dialect,
)
from pandas.core.arrays import (
    ArrowExtensionArray,
    BaseMaskedArray,
    NumpyExtensionArray,
)
from pandas.core.arrays._mixins import NDArrayBackedExtensionArray
from pandas.core.construction import extract_array

if TYPE_CHECKING:
    from pandas._typing import (
        Dtype,
        NpDtype,
    )


UNSIGNED_INT_NUMPY_DTYPES: list[NpDtype] = ["uint8", "uint16", "uint32", "uint64"]
UNSIGNED_INT_EA_DTYPES: list[Dtype] = ["UInt8", "UInt16", "UInt32", "UInt64"]
SIGNED_INT_NUMPY_DTYPES: list[NpDtype] = [int, "int8", "int16", "int32", "int64"]
SIGNED_INT_EA_DTYPES: list[Dtype] = ["Int8", "Int16", "Int32", "Int64"]
ALL_INT_NUMPY_DTYPES = UNSIGNED_INT_NUMPY_DTYPES + SIGNED_INT_NUMPY_DTYPES
ALL_INT_EA_DTYPES = UNSIGNED_INT_EA_DTYPES + SIGNED_INT_EA_DTYPES
ALL_INT_DTYPES: list[Dtype] = [*ALL_INT_NUMPY_DTYPES, *ALL_INT_EA_DTYPES]

FLOAT_NUMPY_DTYPES: list[NpDtype] = [float, "float32", "float64"]
FLOAT_EA_DTYPES: list[Dtype] = ["Float32", "Float64"]
ALL_FLOAT_DTYPES: list[Dtype] = [*FLOAT_NUMPY_DTYPES, *FLOAT_EA_DTYPES]

COMPLEX_DTYPES: list[Dtype] = [complex, "complex64", "complex128"]
if using_string_dtype():
    STRING_DTYPES: list[Dtype] = ["U"]
else:
    STRING_DTYPES: list[Dtype] = [str, "str", "U"]  # type: ignore[no-redef]
COMPLEX_FLOAT_DTYPES: list[Dtype] = [*COMPLEX_DTYPES, *FLOAT_NUMPY_DTYPES]

DATETIME64_DTYPES: list[Dtype] = ["datetime64[ns]", "M8[ns]"]
TIMEDELTA64_DTYPES: list[Dtype] = ["timedelta64[ns]", "m8[ns]"]

BOOL_DTYPES: list[Dtype] = [bool, "bool"]
BYTES_DTYPES: list[Dtype] = [bytes, "bytes"]
OBJECT_DTYPES: list[Dtype] = [object, "object"]

ALL_REAL_NUMPY_DTYPES = FLOAT_NUMPY_DTYPES + ALL_INT_NUMPY_DTYPES
ALL_REAL_EXTENSION_DTYPES = FLOAT_EA_DTYPES + ALL_INT_EA_DTYPES
ALL_REAL_DTYPES: list[Dtype] = [*ALL_REAL_NUMPY_DTYPES, *ALL_REAL_EXTENSION_DTYPES]
ALL_NUMERIC_DTYPES: list[Dtype] = [*ALL_REAL_DTYPES, *COMPLEX_DTYPES]

ALL_NUMPY_DTYPES = (
    ALL_REAL_NUMPY_DTYPES
    + COMPLEX_DTYPES
    + STRING_DTYPES
    + DATETIME64_DTYPES
    + TIMEDELTA64_DTYPES
    + BOOL_DTYPES
    + OBJECT_DTYPES
    + BYTES_DTYPES
)

NARROW_NP_DTYPES = [
    np.float16,
    np.float32,
    np.int8,
    np.int16,
    np.int32,
    np.uint8,
    np.uint16,
    np.uint32,
]

PYTHON_DATA_TYPES = [
    str,
    int,
    float,
    complex,
    list,
    tuple,
    range,
    dict,
    set,
    frozenset,
    bool,
    bytes,
    bytearray,
    memoryview,
]

ENDIAN = {"little": "<", "big": ">"}[byteorder]

NULL_OBJECTS = [None, np.nan, pd.NaT, float("nan"), pd.NA, Decimal("NaN")]
NP_NAT_OBJECTS = [
    cls("NaT", unit)
    for cls in [np.datetime64, np.timedelta64]
    for unit in [
        "Y",
        "M",
        "W",
        "D",
        "h",
        "m",
        "s",
        "ms",
        "us",
        "ns",
        "ps",
        "fs",
        "as",
    ]
]

if not pa_version_under10p1:
    import pyarrow as pa

    UNSIGNED_INT_PYARROW_DTYPES = [pa.uint8(), pa.uint16(), pa.uint32(), pa.uint64()]
    SIGNED_INT_PYARROW_DTYPES = [pa.int8(), pa.int16(), pa.int32(), pa.int64()]
    ALL_INT_PYARROW_DTYPES = UNSIGNED_INT_PYARROW_DTYPES + SIGNED_INT_PYARROW_DTYPES
    ALL_INT_PYARROW_DTYPES_STR_REPR = [
        str(ArrowDtype(typ)) for typ in ALL_INT_PYARROW_DTYPES
    ]

    # pa.float16 doesn't seem supported
    # https://github.com/apache/arrow/blob/master/python/pyarrow/src/arrow/python/helpers.cc#L86
    FLOAT_PYARROW_DTYPES = [pa.float32(), pa.float64()]
    FLOAT_PYARROW_DTYPES_STR_REPR = [
        str(ArrowDtype(typ)) for typ in FLOAT_PYARROW_DTYPES
    ]
    DECIMAL_PYARROW_DTYPES = [pa.decimal128(7, 3)]
    STRING_PYARROW_DTYPES = [pa.string()]
    BINARY_PYARROW_DTYPES = [pa.binary()]

    TIME_PYARROW_DTYPES = [
        pa.time32("s"),
        pa.time32("ms"),
        pa.time64("us"),
        pa.time64("ns"),
    ]
    DATE_PYARROW_DTYPES = [pa.date32(), pa.date64()]
    DATETIME_PYARROW_DTYPES = [
        pa.timestamp(unit=unit, tz=tz)
        for unit in ["s", "ms", "us", "ns"]
        for tz in [None, "UTC", "US/Pacific", "US/Eastern"]
    ]
    TIMEDELTA_PYARROW_DTYPES = [pa.duration(unit) for unit in ["s", "ms", "us", "ns"]]

    BOOL_PYARROW_DTYPES = [pa.bool_()]

    # TODO: Add container like pyarrow types:
    #  https://arrow.apache.org/docs/python/api/datatypes.html#factory-functions
    ALL_PYARROW_DTYPES = (
        ALL_INT_PYARROW_DTYPES
        + FLOAT_PYARROW_DTYPES
        + DECIMAL_PYARROW_DTYPES
        + STRING_PYARROW_DTYPES
        + BINARY_PYARROW_DTYPES
        + TIME_PYARROW_DTYPES
        + DATE_PYARROW_DTYPES
        + DATETIME_PYARROW_DTYPES
        + TIMEDELTA_PYARROW_DTYPES
        + BOOL_PYARROW_DTYPES
    )
    ALL_REAL_PYARROW_DTYPES_STR_REPR = (
        ALL_INT_PYARROW_DTYPES_STR_REPR + FLOAT_PYARROW_DTYPES_STR_REPR
    )
else:
    FLOAT_PYARROW_DTYPES_STR_REPR = []
    ALL_INT_PYARROW_DTYPES_STR_REPR = []
    ALL_PYARROW_DTYPES = []
    ALL_REAL_PYARROW_DTYPES_STR_REPR = []

ALL_REAL_NULLABLE_DTYPES = (
    FLOAT_NUMPY_DTYPES + ALL_REAL_EXTENSION_DTYPES + ALL_REAL_PYARROW_DTYPES_STR_REPR
)

arithmetic_dunder_methods = [
    "__add__",
    "__radd__",
    "__sub__",
    "__rsub__",
    "__mul__",
    "__rmul__",
    "__floordiv__",
    "__rfloordiv__",
    "__truediv__",
    "__rtruediv__",
    "__pow__",
    "__rpow__",
    "__mod__",
    "__rmod__",
]

comparison_dunder_methods = ["__eq__", "__ne__", "__le__", "__lt__", "__ge__", "__gt__"]


# -----------------------------------------------------------------------------
# Comparators


def box_expected(expected, box_cls, transpose: bool = True):
    """
    Helper function to wrap the expected output of a test in a given box_class.

    Parameters
    ----------
    expected : np.ndarray, Index, Series
    box_cls : {Index, Series, DataFrame}

    Returns
    -------
    subclass of box_cls
    """
    if box_cls is pd.array:
        if isinstance(expected, RangeIndex):
            # pd.array would return an IntegerArray
            expected = NumpyExtensionArray(np.asarray(expected._values))
        else:
            expected = pd.array(expected, copy=False)
    elif box_cls is Index:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", "Dtype inference", category=FutureWarning)
            expected = Index(expected)
    elif box_cls is Series:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", "Dtype inference", category=FutureWarning)
            expected = Series(expected)
    elif box_cls is DataFrame:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", "Dtype inference", category=FutureWarning)
            expected = Series(expected).to_frame()
        if transpose:
            # for vector operations, we need a DataFrame to be a single-row,
            #  not a single-column, in order to operate against non-DataFrame
            #  vectors of the same length. But convert to two rows to avoid
            #  single-row special cases in datetime arithmetic
            expected = expected.T
            expected = pd.concat([expected] * 2, ignore_index=True)
    elif box_cls is np.ndarray or box_cls is np.array:
        expected = np.array(expected)
    elif box_cls is to_array:
        expected = to_array(expected)
    else:
        raise NotImplementedError(box_cls)
    return expected


def to_array(obj):
    """
    Similar to pd.array, but does not cast numpy dtypes to nullable dtypes.
    """
    # temporary implementation until we get pd.array in place
    dtype = getattr(obj, "dtype", None)

    if dtype is None:
        return np.asarray(obj)

    return extract_array(obj, extract_numpy=True)


class SubclassedSeries(Series):
    _metadata = ["testattr", "name"]

    @property
    def _constructor(self):
        # For testing, those properties return a generic callable, and not
        # the actual class. In this case that is equivalent, but it is to
        # ensure we don't rely on the property returning a class
        # See https://github.com/pandas-dev/pandas/pull/46018 and
        # https://github.com/pandas-dev/pandas/issues/32638 and linked issues
        return lambda *args, **kwargs: SubclassedSeries(*args, **kwargs)

    @property
    def _constructor_expanddim(self):
        return lambda *args, **kwargs: SubclassedDataFrame(*args, **kwargs)


class SubclassedDataFrame(DataFrame):
    _metadata = ["testattr"]

    @property
    def _constructor(self):
        return lambda *args, **kwargs: SubclassedDataFrame(*args, **kwargs)

    @property
    def _constructor_sliced(self):
        return lambda *args, **kwargs: SubclassedSeries(*args, **kwargs)


def convert_rows_list_to_csv_str(rows_list: list[str]) -> str:
    """
    Convert list of CSV rows to single CSV-formatted string for current OS.

    This method is used for creating expected value of to_csv() method.

    Parameters
    ----------
    rows_list : List[str]
        Each element represents the row of csv.

    Returns
    -------
    str
        Expected output of to_csv() in current OS.
    """
    sep = os.linesep
    return sep.join(rows_list) + sep


def external_error_raised(expected_exception: type[Exception]) -> ContextManager:
    """
    Helper function to mark pytest.raises that have an external error message.

    Parameters
    ----------
    expected_exception : Exception
        Expected error to raise.

    Returns
    -------
    Callable
        Regular `pytest.raises` function with `match` equal to `None`.
    """
    import pytest

    return pytest.raises(expected_exception, match=None)


cython_table = pd.core.common._cython_table.items()


def get_cython_table_params(ndframe, func_names_and_expected):
    """
    Combine frame, functions from com._cython_table
    keys and expected result.

    Parameters
    ----------
    ndframe : DataFrame or Series
    func_names_and_expected : Sequence of two items
        The first item is a name of a NDFrame method ('sum', 'prod') etc.
        The second item is the expected return value.

    Returns
    -------
    list
        List of three items (DataFrame, function, expected result)
    """
    results = []
    for func_name, expected in func_names_and_expected:
        results.append((ndframe, func_name, expected))
        results += [
            (ndframe, func, expected)
            for func, name in cython_table
            if name == func_name
        ]
    return results


def get_op_from_name(op_name: str) -> Callable:
    """
    The operator function for a given op name.

    Parameters
    ----------
    op_name : str
        The op name, in form of "add" or "__add__".

    Returns
    -------
    function
        A function performing the operation.
    """
    short_opname = op_name.strip("_")
    try:
        op = getattr(operator, short_opname)
    except AttributeError:
        # Assume it is the reverse operator
        rop = getattr(operator, short_opname[1:])
        op = lambda x, y: rop(y, x)

    return op


# -----------------------------------------------------------------------------
# Indexing test helpers


def getitem(x):
    return x


def setitem(x):
    return x


def loc(x):
    return x.loc


def iloc(x):
    return x.iloc


def at(x):
    return x.at


def iat(x):
    return x.iat


# -----------------------------------------------------------------------------

_UNITS = ["s", "ms", "us", "ns"]


def get_finest_unit(left: str, right: str):
    """
    Find the higher of two datetime64 units.
    """
    if _UNITS.index(left) >= _UNITS.index(right):
        return left
    return right


def shares_memory(left, right) -> bool:
    """
    Pandas-compat for np.shares_memory.
    """
    if isinstance(left, np.ndarray) and isinstance(right, np.ndarray):
        return np.shares_memory(left, right)
    elif isinstance(left, np.ndarray):
        # Call with reversed args to get to unpacking logic below.
        return shares_memory(right, left)

    if isinstance(left, RangeIndex):
        return False
    if isinstance(left, MultiIndex):
        return shares_memory(left._codes, right)
    if isinstance(left, (Index, Series)):
        if isinstance(right, (Index, Series)):
            return shares_memory(left._values, right._values)
        return shares_memory(left._values, right)

    if isinstance(left, NDArrayBackedExtensionArray):
        return shares_memory(left._ndarray, right)
    if isinstance(left, pd.core.arrays.SparseArray):
        return shares_memory(left.sp_values, right)
    if isinstance(left, pd.core.arrays.IntervalArray):
        return shares_memory(left._left, right) or shares_memory(left._right, right)

    if isinstance(left, ArrowExtensionArray):
        if isinstance(right, ArrowExtensionArray):
            # https://github.com/pandas-dev/pandas/pull/43930#discussion_r736862669
            left_pa_data = left._pa_array
            right_pa_data = right._pa_array
            left_buf1 = left_pa_data.chunk(0).buffers()[1]
            right_buf1 = right_pa_data.chunk(0).buffers()[1]
            return left_buf1.address == right_buf1.address
        else:
            # if we have one one ArrowExtensionArray and one other array, assume
            # they can only share memory if they share the same numpy buffer
            return np.shares_memory(left, right)

    if isinstance(left, BaseMaskedArray) and isinstance(right, BaseMaskedArray):
        # By convention, we'll say these share memory if they share *either*
        #  the _data or the _mask
        return np.shares_memory(left._data, right._data) or np.shares_memory(
            left._mask, right._mask
        )

    if isinstance(left, DataFrame) and len(left._mgr.arrays) == 1:
        arr = left._mgr.arrays[0]
        return shares_memory(arr, right)

    raise NotImplementedError(type(left), type(right))


__all__ = [
    "ALL_INT_EA_DTYPES",
    "ALL_INT_NUMPY_DTYPES",
    "ALL_NUMPY_DTYPES",
    "ALL_REAL_NUMPY_DTYPES",
    "assert_almost_equal",
    "assert_attr_equal",
    "assert_categorical_equal",
    "assert_class_equal",
    "assert_contains_all",
    "assert_copy",
    "assert_datetime_array_equal",
    "assert_dict_equal",
    "assert_equal",
    "assert_extension_array_equal",
    "assert_frame_equal",
    "assert_index_equal",
    "assert_indexing_slices_equivalent",
    "assert_interval_array_equal",
    "assert_is_sorted",
    "assert_is_valid_plot_return_object",
    "assert_metadata_equivalent",
    "assert_numpy_array_equal",
    "assert_period_array_equal",
    "assert_produces_warning",
    "assert_series_equal",
    "assert_sp_array_equal",
    "assert_timedelta_array_equal",
    "assert_cow_warning",
    "at",
    "BOOL_DTYPES",
    "box_expected",
    "BYTES_DTYPES",
    "can_set_locale",
    "COMPLEX_DTYPES",
    "convert_rows_list_to_csv_str",
    "DATETIME64_DTYPES",
    "decompress_file",
    "ENDIAN",
    "ensure_clean",
    "external_error_raised",
    "FLOAT_EA_DTYPES",
    "FLOAT_NUMPY_DTYPES",
    "get_cython_table_params",
    "get_dtype",
    "getitem",
    "get_locales",
    "get_finest_unit",
    "get_obj",
    "get_op_from_name",
    "iat",
    "iloc",
    "loc",
    "maybe_produces_warning",
    "NARROW_NP_DTYPES",
    "NP_NAT_OBJECTS",
    "NULL_OBJECTS",
    "OBJECT_DTYPES",
    "raise_assert_detail",
    "raises_chained_assignment_error",
    "round_trip_localpath",
    "round_trip_pathlib",
    "round_trip_pickle",
    "setitem",
    "set_locale",
    "set_timezone",
    "shares_memory",
    "SIGNED_INT_EA_DTYPES",
    "SIGNED_INT_NUMPY_DTYPES",
    "STRING_DTYPES",
    "SubclassedDataFrame",
    "SubclassedSeries",
    "TIMEDELTA64_DTYPES",
    "to_array",
    "UNSIGNED_INT_EA_DTYPES",
    "UNSIGNED_INT_NUMPY_DTYPES",
    "use_numexpr",
    "with_csv_dialect",
    "write_to_compressed",
]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\vector\point.py ===
from .vector import Vector, _check_vector
from .frame import _check_frame
from warnings import warn
from sympy.utilities.misc import filldedent

__all__ = ['Point']


class Point:
    """This object represents a point in a dynamic system.

    It stores the: position, velocity, and acceleration of a point.
    The position is a vector defined as the vector distance from a parent
    point to this point.

    Parameters
    ==========

    name : string
        The display name of the Point

    Examples
    ========

    >>> from sympy.physics.vector import Point, ReferenceFrame, dynamicsymbols
    >>> from sympy.physics.vector import init_vprinting
    >>> init_vprinting(pretty_print=False)
    >>> N = ReferenceFrame('N')
    >>> O = Point('O')
    >>> P = Point('P')
    >>> u1, u2, u3 = dynamicsymbols('u1 u2 u3')
    >>> O.set_vel(N, u1 * N.x + u2 * N.y + u3 * N.z)
    >>> O.acc(N)
    u1'*N.x + u2'*N.y + u3'*N.z

    ``symbols()`` can be used to create multiple Points in a single step, for
    example:

    >>> from sympy.physics.vector import Point, ReferenceFrame, dynamicsymbols
    >>> from sympy.physics.vector import init_vprinting
    >>> init_vprinting(pretty_print=False)
    >>> from sympy import symbols
    >>> N = ReferenceFrame('N')
    >>> u1, u2 = dynamicsymbols('u1 u2')
    >>> A, B = symbols('A B', cls=Point)
    >>> type(A)
    <class 'sympy.physics.vector.point.Point'>
    >>> A.set_vel(N, u1 * N.x + u2 * N.y)
    >>> B.set_vel(N, u2 * N.x + u1 * N.y)
    >>> A.acc(N) - B.acc(N)
    (u1' - u2')*N.x + (-u1' + u2')*N.y

    """

    def __init__(self, name):
        """Initialization of a Point object. """
        self.name = name
        self._pos_dict = {}
        self._vel_dict = {}
        self._acc_dict = {}
        self._pdlist = [self._pos_dict, self._vel_dict, self._acc_dict]

    def __str__(self):
        return self.name

    __repr__ = __str__

    def _check_point(self, other):
        if not isinstance(other, Point):
            raise TypeError('A Point must be supplied')

    def _pdict_list(self, other, num):
        """Returns a list of points that gives the shortest path with respect
        to position, velocity, or acceleration from this point to the provided
        point.

        Parameters
        ==========
        other : Point
            A point that may be related to this point by position, velocity, or
            acceleration.
        num : integer
            0 for searching the position tree, 1 for searching the velocity
            tree, and 2 for searching the acceleration tree.

        Returns
        =======
        list of Points
            A sequence of points from self to other.

        Notes
        =====

        It is not clear if num = 1 or num = 2 actually works because the keys
        to ``_vel_dict`` and ``_acc_dict`` are :class:`ReferenceFrame` objects
        which do not have the ``_pdlist`` attribute.

        """
        outlist = [[self]]
        oldlist = [[]]
        while outlist != oldlist:
            oldlist = outlist.copy()
            for v in outlist:
                templist = v[-1]._pdlist[num].keys()
                for v2 in templist:
                    if not v.__contains__(v2):
                        littletemplist = v + [v2]
                        if not outlist.__contains__(littletemplist):
                            outlist.append(littletemplist)
        for v in oldlist:
            if v[-1] != other:
                outlist.remove(v)
        outlist.sort(key=len)
        if len(outlist) != 0:
            return outlist[0]
        raise ValueError('No Connecting Path found between ' + other.name +
                         ' and ' + self.name)

    def a1pt_theory(self, otherpoint, outframe, interframe):
        """Sets the acceleration of this point with the 1-point theory.

        The 1-point theory for point acceleration looks like this:

        ^N a^P = ^B a^P + ^N a^O + ^N alpha^B x r^OP + ^N omega^B x (^N omega^B
        x r^OP) + 2 ^N omega^B x ^B v^P

        where O is a point fixed in B, P is a point moving in B, and B is
        rotating in frame N.

        Parameters
        ==========

        otherpoint : Point
            The first point of the 1-point theory (O)
        outframe : ReferenceFrame
            The frame we want this point's acceleration defined in (N)
        fixedframe : ReferenceFrame
            The intermediate frame in this calculation (B)

        Examples
        ========

        >>> from sympy.physics.vector import Point, ReferenceFrame
        >>> from sympy.physics.vector import dynamicsymbols
        >>> from sympy.physics.vector import init_vprinting
        >>> init_vprinting(pretty_print=False)
        >>> q = dynamicsymbols('q')
        >>> q2 = dynamicsymbols('q2')
        >>> qd = dynamicsymbols('q', 1)
        >>> q2d = dynamicsymbols('q2', 1)
        >>> N = ReferenceFrame('N')
        >>> B = ReferenceFrame('B')
        >>> B.set_ang_vel(N, 5 * B.y)
        >>> O = Point('O')
        >>> P = O.locatenew('P', q * B.x + q2 * B.y)
        >>> P.set_vel(B, qd * B.x + q2d * B.y)
        >>> O.set_vel(N, 0)
        >>> P.a1pt_theory(O, N, B)
        (-25*q + q'')*B.x + q2''*B.y - 10*q'*B.z

        """

        _check_frame(outframe)
        _check_frame(interframe)
        self._check_point(otherpoint)
        dist = self.pos_from(otherpoint)
        v = self.vel(interframe)
        a1 = otherpoint.acc(outframe)
        a2 = self.acc(interframe)
        omega = interframe.ang_vel_in(outframe)
        alpha = interframe.ang_acc_in(outframe)
        self.set_acc(outframe, a2 + 2 * (omega.cross(v)) + a1 +
                     (alpha.cross(dist)) + (omega.cross(omega.cross(dist))))
        return self.acc(outframe)

    def a2pt_theory(self, otherpoint, outframe, fixedframe):
        """Sets the acceleration of this point with the 2-point theory.

        The 2-point theory for point acceleration looks like this:

        ^N a^P = ^N a^O + ^N alpha^B x r^OP + ^N omega^B x (^N omega^B x r^OP)

        where O and P are both points fixed in frame B, which is rotating in
        frame N.

        Parameters
        ==========

        otherpoint : Point
            The first point of the 2-point theory (O)
        outframe : ReferenceFrame
            The frame we want this point's acceleration defined in (N)
        fixedframe : ReferenceFrame
            The frame in which both points are fixed (B)

        Examples
        ========

        >>> from sympy.physics.vector import Point, ReferenceFrame, dynamicsymbols
        >>> from sympy.physics.vector import init_vprinting
        >>> init_vprinting(pretty_print=False)
        >>> q = dynamicsymbols('q')
        >>> qd = dynamicsymbols('q', 1)
        >>> N = ReferenceFrame('N')
        >>> B = N.orientnew('B', 'Axis', [q, N.z])
        >>> O = Point('O')
        >>> P = O.locatenew('P', 10 * B.x)
        >>> O.set_vel(N, 5 * N.x)
        >>> P.a2pt_theory(O, N, B)
        - 10*q'**2*B.x + 10*q''*B.y

        """

        _check_frame(outframe)
        _check_frame(fixedframe)
        self._check_point(otherpoint)
        dist = self.pos_from(otherpoint)
        a = otherpoint.acc(outframe)
        omega = fixedframe.ang_vel_in(outframe)
        alpha = fixedframe.ang_acc_in(outframe)
        self.set_acc(outframe, a + (alpha.cross(dist)) +
                     (omega.cross(omega.cross(dist))))
        return self.acc(outframe)

    def acc(self, frame):
        """The acceleration Vector of this Point in a ReferenceFrame.

        Parameters
        ==========

        frame : ReferenceFrame
            The frame in which the returned acceleration vector will be defined
            in.

        Examples
        ========

        >>> from sympy.physics.vector import Point, ReferenceFrame
        >>> N = ReferenceFrame('N')
        >>> p1 = Point('p1')
        >>> p1.set_acc(N, 10 * N.x)
        >>> p1.acc(N)
        10*N.x

        """

        _check_frame(frame)
        if not (frame in self._acc_dict):
            if self.vel(frame) != 0:
                return (self._vel_dict[frame]).dt(frame)
            else:
                return Vector(0)
        return self._acc_dict[frame]

    def locatenew(self, name, value):
        """Creates a new point with a position defined from this point.

        Parameters
        ==========

        name : str
            The name for the new point
        value : Vector
            The position of the new point relative to this point

        Examples
        ========

        >>> from sympy.physics.vector import ReferenceFrame, Point
        >>> N = ReferenceFrame('N')
        >>> P1 = Point('P1')
        >>> P2 = P1.locatenew('P2', 10 * N.x)

        """

        if not isinstance(name, str):
            raise TypeError('Must supply a valid name')
        if value == 0:
            value = Vector(0)
        value = _check_vector(value)
        p = Point(name)
        p.set_pos(self, value)
        self.set_pos(p, -value)
        return p

    def pos_from(self, otherpoint):
        """Returns a Vector distance between this Point and the other Point.

        Parameters
        ==========

        otherpoint : Point
            The otherpoint we are locating this one relative to

        Examples
        ========

        >>> from sympy.physics.vector import Point, ReferenceFrame
        >>> N = ReferenceFrame('N')
        >>> p1 = Point('p1')
        >>> p2 = Point('p2')
        >>> p1.set_pos(p2, 10 * N.x)
        >>> p1.pos_from(p2)
        10*N.x

        """

        outvec = Vector(0)
        plist = self._pdict_list(otherpoint, 0)
        for i in range(len(plist) - 1):
            outvec += plist[i]._pos_dict[plist[i + 1]]
        return outvec

    def set_acc(self, frame, value):
        """Used to set the acceleration of this Point in a ReferenceFrame.

        Parameters
        ==========

        frame : ReferenceFrame
            The frame in which this point's acceleration is defined
        value : Vector
            The vector value of this point's acceleration in the frame

        Examples
        ========

        >>> from sympy.physics.vector import Point, ReferenceFrame
        >>> N = ReferenceFrame('N')
        >>> p1 = Point('p1')
        >>> p1.set_acc(N, 10 * N.x)
        >>> p1.acc(N)
        10*N.x

        """

        if value == 0:
            value = Vector(0)
        value = _check_vector(value)
        _check_frame(frame)
        self._acc_dict.update({frame: value})

    def set_pos(self, otherpoint, value):
        """Used to set the position of this point w.r.t. another point.

        Parameters
        ==========

        otherpoint : Point
            The other point which this point's location is defined relative to
        value : Vector
            The vector which defines the location of this point

        Examples
        ========

        >>> from sympy.physics.vector import Point, ReferenceFrame
        >>> N = ReferenceFrame('N')
        >>> p1 = Point('p1')
        >>> p2 = Point('p2')
        >>> p1.set_pos(p2, 10 * N.x)
        >>> p1.pos_from(p2)
        10*N.x

        """

        if value == 0:
            value = Vector(0)
        value = _check_vector(value)
        self._check_point(otherpoint)
        self._pos_dict.update({otherpoint: value})
        otherpoint._pos_dict.update({self: -value})

    def set_vel(self, frame, value):
        """Sets the velocity Vector of this Point in a ReferenceFrame.

        Parameters
        ==========

        frame : ReferenceFrame
            The frame in which this point's velocity is defined
        value : Vector
            The vector value of this point's velocity in the frame

        Examples
        ========

        >>> from sympy.physics.vector import Point, ReferenceFrame
        >>> N = ReferenceFrame('N')
        >>> p1 = Point('p1')
        >>> p1.set_vel(N, 10 * N.x)
        >>> p1.vel(N)
        10*N.x

        """

        if value == 0:
            value = Vector(0)
        value = _check_vector(value)
        _check_frame(frame)
        self._vel_dict.update({frame: value})

    def v1pt_theory(self, otherpoint, outframe, interframe):
        """Sets the velocity of this point with the 1-point theory.

        The 1-point theory for point velocity looks like this:

        ^N v^P = ^B v^P + ^N v^O + ^N omega^B x r^OP

        where O is a point fixed in B, P is a point moving in B, and B is
        rotating in frame N.

        Parameters
        ==========

        otherpoint : Point
            The first point of the 1-point theory (O)
        outframe : ReferenceFrame
            The frame we want this point's velocity defined in (N)
        interframe : ReferenceFrame
            The intermediate frame in this calculation (B)

        Examples
        ========

        >>> from sympy.physics.vector import Point, ReferenceFrame
        >>> from sympy.physics.vector import dynamicsymbols
        >>> from sympy.physics.vector import init_vprinting
        >>> init_vprinting(pretty_print=False)
        >>> q = dynamicsymbols('q')
        >>> q2 = dynamicsymbols('q2')
        >>> qd = dynamicsymbols('q', 1)
        >>> q2d = dynamicsymbols('q2', 1)
        >>> N = ReferenceFrame('N')
        >>> B = ReferenceFrame('B')
        >>> B.set_ang_vel(N, 5 * B.y)
        >>> O = Point('O')
        >>> P = O.locatenew('P', q * B.x + q2 * B.y)
        >>> P.set_vel(B, qd * B.x + q2d * B.y)
        >>> O.set_vel(N, 0)
        >>> P.v1pt_theory(O, N, B)
        q'*B.x + q2'*B.y - 5*q*B.z

        """

        _check_frame(outframe)
        _check_frame(interframe)
        self._check_point(otherpoint)
        dist = self.pos_from(otherpoint)
        v1 = self.vel(interframe)
        v2 = otherpoint.vel(outframe)
        omega = interframe.ang_vel_in(outframe)
        self.set_vel(outframe, v1 + v2 + (omega.cross(dist)))
        return self.vel(outframe)

    def v2pt_theory(self, otherpoint, outframe, fixedframe):
        """Sets the velocity of this point with the 2-point theory.

        The 2-point theory for point velocity looks like this:

        ^N v^P = ^N v^O + ^N omega^B x r^OP

        where O and P are both points fixed in frame B, which is rotating in
        frame N.

        Parameters
        ==========

        otherpoint : Point
            The first point of the 2-point theory (O)
        outframe : ReferenceFrame
            The frame we want this point's velocity defined in (N)
        fixedframe : ReferenceFrame
            The frame in which both points are fixed (B)

        Examples
        ========

        >>> from sympy.physics.vector import Point, ReferenceFrame, dynamicsymbols
        >>> from sympy.physics.vector import init_vprinting
        >>> init_vprinting(pretty_print=False)
        >>> q = dynamicsymbols('q')
        >>> qd = dynamicsymbols('q', 1)
        >>> N = ReferenceFrame('N')
        >>> B = N.orientnew('B', 'Axis', [q, N.z])
        >>> O = Point('O')
        >>> P = O.locatenew('P', 10 * B.x)
        >>> O.set_vel(N, 5 * N.x)
        >>> P.v2pt_theory(O, N, B)
        5*N.x + 10*q'*B.y

        """

        _check_frame(outframe)
        _check_frame(fixedframe)
        self._check_point(otherpoint)
        dist = self.pos_from(otherpoint)
        v = otherpoint.vel(outframe)
        omega = fixedframe.ang_vel_in(outframe)
        self.set_vel(outframe, v + (omega.cross(dist)))
        return self.vel(outframe)

    def vel(self, frame):
        """The velocity Vector of this Point in the ReferenceFrame.

        Parameters
        ==========

        frame : ReferenceFrame
            The frame in which the returned velocity vector will be defined in

        Examples
        ========

        >>> from sympy.physics.vector import Point, ReferenceFrame, dynamicsymbols
        >>> N = ReferenceFrame('N')
        >>> p1 = Point('p1')
        >>> p1.set_vel(N, 10 * N.x)
        >>> p1.vel(N)
        10*N.x

        Velocities will be automatically calculated if possible, otherwise a
        ``ValueError`` will be returned. If it is possible to calculate
        multiple different velocities from the relative points, the points
        defined most directly relative to this point will be used. In the case
        of inconsistent relative positions of points, incorrect velocities may
        be returned. It is up to the user to define prior relative positions
        and velocities of points in a self-consistent way.

        >>> p = Point('p')
        >>> q = dynamicsymbols('q')
        >>> p.set_vel(N, 10 * N.x)
        >>> p2 = Point('p2')
        >>> p2.set_pos(p, q*N.x)
        >>> p2.vel(N)
        (Derivative(q(t), t) + 10)*N.x

        """

        _check_frame(frame)
        if not (frame in self._vel_dict):
            valid_neighbor_found = False
            is_cyclic = False
            visited = []
            queue = [self]
            candidate_neighbor = []
            while queue:  # BFS to find nearest point
                node = queue.pop(0)
                if node not in visited:
                    visited.append(node)
                    for neighbor, neighbor_pos in node._pos_dict.items():
                        if neighbor in visited:
                            continue
                        try:
                            # Checks if pos vector is valid
                            neighbor_pos.express(frame)
                        except ValueError:
                            continue
                        if neighbor in queue:
                            is_cyclic = True
                        try:
                            # Checks if point has its vel defined in req frame
                            neighbor_velocity = neighbor._vel_dict[frame]
                        except KeyError:
                            queue.append(neighbor)
                            continue
                        candidate_neighbor.append(neighbor)
                        if not valid_neighbor_found:
                            self.set_vel(frame, self.pos_from(neighbor).dt(frame) + neighbor_velocity)
                            valid_neighbor_found = True
            if is_cyclic:
                warn(filldedent("""
                Kinematic loops are defined among the positions of points. This
                is likely not desired and may cause errors in your calculations.
                """))
            if len(candidate_neighbor) > 1:
                warn(filldedent(f"""
                Velocity of {self.name} automatically calculated based on point
                {candidate_neighbor[0].name} but it is also possible from
                points(s): {str(candidate_neighbor[1:])}. Velocities from these
                points are not necessarily the same. This may cause errors in
                your calculations."""))
            if valid_neighbor_found:
                return self._vel_dict[frame]
            else:
                raise ValueError(filldedent(f"""
                Velocity of point {self.name} has not been defined in
                ReferenceFrame {frame.name}."""))

        return self._vel_dict[frame]

    def partial_velocity(self, frame, *gen_speeds):
        """Returns the partial velocities of the linear velocity vector of this
        point in the given frame with respect to one or more provided
        generalized speeds.

        Parameters
        ==========
        frame : ReferenceFrame
            The frame with which the velocity is defined in.
        gen_speeds : functions of time
            The generalized speeds.

        Returns
        =======
        partial_velocities : tuple of Vector
            The partial velocity vectors corresponding to the provided
            generalized speeds.

        Examples
        ========

        >>> from sympy.physics.vector import ReferenceFrame, Point
        >>> from sympy.physics.vector import dynamicsymbols
        >>> N = ReferenceFrame('N')
        >>> A = ReferenceFrame('A')
        >>> p = Point('p')
        >>> u1, u2 = dynamicsymbols('u1, u2')
        >>> p.set_vel(N, u1 * N.x + u2 * A.y)
        >>> p.partial_velocity(N, u1)
        N.x
        >>> p.partial_velocity(N, u1, u2)
        (N.x, A.y)

        """

        from sympy.physics.vector.functions import partial_velocity

        vel = self.vel(frame)
        partials = partial_velocity([vel], gen_speeds, frame)[0]

        if len(partials) == 1:
            return partials[0]
        else:
            return tuple(partials)