
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_core\_windows_cffi.py ===
from __future__ import annotations

import enum
import re
from typing import TYPE_CHECKING, NewType, NoReturn, Protocol, cast

if TYPE_CHECKING:
    from typing_extensions import TypeAlias

import cffi

################################################################
# Functions and types
################################################################

LIB = """
// https://msdn.microsoft.com/en-us/library/windows/desktop/aa383751(v=vs.85).aspx
typedef int BOOL;
typedef unsigned char BYTE;
typedef BYTE BOOLEAN;
typedef void* PVOID;
typedef PVOID HANDLE;
typedef unsigned long DWORD;
typedef unsigned long ULONG;
typedef unsigned int NTSTATUS;
typedef unsigned long u_long;
typedef ULONG *PULONG;
typedef const void *LPCVOID;
typedef void *LPVOID;
typedef const wchar_t *LPCWSTR;

typedef uintptr_t ULONG_PTR;
typedef uintptr_t UINT_PTR;

typedef UINT_PTR SOCKET;

typedef struct _OVERLAPPED {
    ULONG_PTR Internal;
    ULONG_PTR InternalHigh;
    union {
        struct {
            DWORD Offset;
            DWORD OffsetHigh;
        } DUMMYSTRUCTNAME;
        PVOID Pointer;
    } DUMMYUNIONNAME;

    HANDLE  hEvent;
} OVERLAPPED, *LPOVERLAPPED;

typedef OVERLAPPED WSAOVERLAPPED;
typedef LPOVERLAPPED LPWSAOVERLAPPED;
typedef PVOID LPSECURITY_ATTRIBUTES;
typedef PVOID LPCSTR;

typedef struct _OVERLAPPED_ENTRY {
    ULONG_PTR lpCompletionKey;
    LPOVERLAPPED lpOverlapped;
    ULONG_PTR Internal;
    DWORD dwNumberOfBytesTransferred;
} OVERLAPPED_ENTRY, *LPOVERLAPPED_ENTRY;

// kernel32.dll
HANDLE WINAPI CreateIoCompletionPort(
  _In_     HANDLE    FileHandle,
  _In_opt_ HANDLE    ExistingCompletionPort,
  _In_     ULONG_PTR CompletionKey,
  _In_     DWORD     NumberOfConcurrentThreads
);

BOOL SetFileCompletionNotificationModes(
  HANDLE FileHandle,
  UCHAR  Flags
);

HANDLE CreateFileW(
  LPCWSTR               lpFileName,
  DWORD                 dwDesiredAccess,
  DWORD                 dwShareMode,
  LPSECURITY_ATTRIBUTES lpSecurityAttributes,
  DWORD                 dwCreationDisposition,
  DWORD                 dwFlagsAndAttributes,
  HANDLE                hTemplateFile
);

BOOL WINAPI CloseHandle(
  _In_ HANDLE hObject
);

BOOL WINAPI PostQueuedCompletionStatus(
  _In_     HANDLE       CompletionPort,
  _In_     DWORD        dwNumberOfBytesTransferred,
  _In_     ULONG_PTR    dwCompletionKey,
  _In_opt_ LPOVERLAPPED lpOverlapped
);

BOOL WINAPI GetQueuedCompletionStatusEx(
  _In_  HANDLE             CompletionPort,
  _Out_ LPOVERLAPPED_ENTRY lpCompletionPortEntries,
  _In_  ULONG              ulCount,
  _Out_ PULONG             ulNumEntriesRemoved,
  _In_  DWORD              dwMilliseconds,
  _In_  BOOL               fAlertable
);

BOOL WINAPI CancelIoEx(
  _In_     HANDLE       hFile,
  _In_opt_ LPOVERLAPPED lpOverlapped
);

BOOL WriteFile(
  HANDLE       hFile,
  LPCVOID      lpBuffer,
  DWORD        nNumberOfBytesToWrite,
  LPDWORD      lpNumberOfBytesWritten,
  LPOVERLAPPED lpOverlapped
);

BOOL ReadFile(
  HANDLE       hFile,
  LPVOID       lpBuffer,
  DWORD        nNumberOfBytesToRead,
  LPDWORD      lpNumberOfBytesRead,
  LPOVERLAPPED lpOverlapped
);

BOOL WINAPI SetConsoleCtrlHandler(
  _In_opt_ void*            HandlerRoutine,
  _In_     BOOL             Add
);

HANDLE CreateEventA(
  LPSECURITY_ATTRIBUTES lpEventAttributes,
  BOOL                  bManualReset,
  BOOL                  bInitialState,
  LPCSTR                lpName
);

BOOL SetEvent(
  HANDLE hEvent
);

BOOL ResetEvent(
  HANDLE hEvent
);

DWORD WaitForSingleObject(
  HANDLE hHandle,
  DWORD  dwMilliseconds
);

DWORD WaitForMultipleObjects(
  DWORD        nCount,
  HANDLE       *lpHandles,
  BOOL         bWaitAll,
  DWORD        dwMilliseconds
);

ULONG RtlNtStatusToDosError(
  NTSTATUS Status
);

int WSAIoctl(
  SOCKET                             s,
  DWORD                              dwIoControlCode,
  LPVOID                             lpvInBuffer,
  DWORD                              cbInBuffer,
  LPVOID                             lpvOutBuffer,
  DWORD                              cbOutBuffer,
  LPDWORD                            lpcbBytesReturned,
  LPWSAOVERLAPPED                    lpOverlapped,
  // actually LPWSAOVERLAPPED_COMPLETION_ROUTINE
  void* lpCompletionRoutine
);

int WSAGetLastError();

BOOL DeviceIoControl(
  HANDLE       hDevice,
  DWORD        dwIoControlCode,
  LPVOID       lpInBuffer,
  DWORD        nInBufferSize,
  LPVOID       lpOutBuffer,
  DWORD        nOutBufferSize,
  LPDWORD      lpBytesReturned,
  LPOVERLAPPED lpOverlapped
);

// From https://github.com/piscisaureus/wepoll/blob/master/src/afd.h
typedef struct _AFD_POLL_HANDLE_INFO {
  HANDLE Handle;
  ULONG Events;
  NTSTATUS Status;
} AFD_POLL_HANDLE_INFO, *PAFD_POLL_HANDLE_INFO;

// This is really defined as a messy union to allow stuff like
// i.DUMMYSTRUCTNAME.LowPart, but we don't need those complications.
// Under all that it's just an int64.
typedef int64_t LARGE_INTEGER;

typedef struct _AFD_POLL_INFO {
  LARGE_INTEGER Timeout;
  ULONG NumberOfHandles;
  ULONG Exclusive;
  AFD_POLL_HANDLE_INFO Handles[1];
} AFD_POLL_INFO, *PAFD_POLL_INFO;

"""

# cribbed from pywincffi
# programmatically strips out those annotations MSDN likes, like _In_
REGEX_SAL_ANNOTATION = re.compile(
    r"\b(_In_|_Inout_|_Out_|_Outptr_|_Reserved_)(opt_)?\b",
)
LIB = REGEX_SAL_ANNOTATION.sub(" ", LIB)

# Other fixups:
# - get rid of FAR, cffi doesn't like it
LIB = re.sub(r"\bFAR\b", " ", LIB)
# - PASCAL is apparently an alias for __stdcall (on modern compilers - modern
#   being _MSC_VER >= 800)
LIB = re.sub(r"\bPASCAL\b", "__stdcall", LIB)

ffi = cffi.api.FFI()
ffi.cdef(LIB)

CData: TypeAlias = cffi.api.FFI.CData
CType: TypeAlias = cffi.api.FFI.CType
AlwaysNull: TypeAlias = CType  # We currently always pass ffi.NULL here.
Handle = NewType("Handle", CData)
HandleArray = NewType("HandleArray", CData)


class _Kernel32(Protocol):
    """Statically typed version of the kernel32.dll functions we use."""

    def CreateIoCompletionPort(
        self,
        FileHandle: Handle,
        ExistingCompletionPort: CData | AlwaysNull,
        CompletionKey: int,
        NumberOfConcurrentThreads: int,
        /,
    ) -> Handle: ...

    def CreateEventA(
        self,
        lpEventAttributes: AlwaysNull,
        bManualReset: bool,
        bInitialState: bool,
        lpName: AlwaysNull,
        /,
    ) -> Handle: ...

    def SetFileCompletionNotificationModes(
        self,
        handle: Handle,
        flags: CompletionModes,
        /,
    ) -> int: ...

    def PostQueuedCompletionStatus(
        self,
        CompletionPort: Handle,
        dwNumberOfBytesTransferred: int,
        dwCompletionKey: int,
        lpOverlapped: CData | AlwaysNull,
        /,
    ) -> bool: ...

    def CancelIoEx(
        self,
        hFile: Handle,
        lpOverlapped: CData | AlwaysNull,
        /,
    ) -> bool: ...

    def WriteFile(
        self,
        hFile: Handle,
        # not sure about this type
        lpBuffer: CData,
        nNumberOfBytesToWrite: int,
        lpNumberOfBytesWritten: AlwaysNull,
        lpOverlapped: _Overlapped,
        /,
    ) -> bool: ...

    def ReadFile(
        self,
        hFile: Handle,
        # not sure about this type
        lpBuffer: CData,
        nNumberOfBytesToRead: int,
        lpNumberOfBytesRead: AlwaysNull,
        lpOverlapped: _Overlapped,
        /,
    ) -> bool: ...

    def GetQueuedCompletionStatusEx(
        self,
        CompletionPort: Handle,
        lpCompletionPortEntries: CData,
        ulCount: int,
        ulNumEntriesRemoved: CData,
        dwMilliseconds: int,
        fAlertable: bool | int,
        /,
    ) -> CData: ...

    def CreateFileW(
        self,
        lpFileName: CData,
        dwDesiredAccess: FileFlags,
        dwShareMode: FileFlags,
        lpSecurityAttributes: AlwaysNull,
        dwCreationDisposition: FileFlags,
        dwFlagsAndAttributes: FileFlags,
        hTemplateFile: AlwaysNull,
        /,
    ) -> Handle: ...

    def WaitForSingleObject(self, hHandle: Handle, dwMilliseconds: int, /) -> CData: ...

    def WaitForMultipleObjects(
        self,
        nCount: int,
        lpHandles: HandleArray,
        bWaitAll: bool,
        dwMilliseconds: int,
        /,
    ) -> ErrorCodes: ...

    def SetEvent(self, handle: Handle, /) -> None: ...

    def CloseHandle(self, handle: Handle, /) -> bool: ...

    def DeviceIoControl(
        self,
        hDevice: Handle,
        dwIoControlCode: int,
        # this is wrong (it's not always null)
        lpInBuffer: AlwaysNull,
        nInBufferSize: int,
        # this is also wrong
        lpOutBuffer: AlwaysNull,
        nOutBufferSize: int,
        lpBytesReturned: AlwaysNull,
        lpOverlapped: CData,
        /,
    ) -> bool: ...


class _Nt(Protocol):
    """Statically typed version of the dtdll.dll functions we use."""

    def RtlNtStatusToDosError(self, status: int, /) -> ErrorCodes: ...


class _Ws2(Protocol):
    """Statically typed version of the ws2_32.dll functions we use."""

    def WSAGetLastError(self) -> int: ...

    def WSAIoctl(
        self,
        socket: CData,
        dwIoControlCode: WSAIoctls,
        lpvInBuffer: AlwaysNull,
        cbInBuffer: int,
        lpvOutBuffer: CData,
        cbOutBuffer: int,
        lpcbBytesReturned: CData,  # int*
        lpOverlapped: AlwaysNull,
        # actually LPWSAOVERLAPPED_COMPLETION_ROUTINE
        lpCompletionRoutine: AlwaysNull,
        /,
    ) -> int: ...


class _DummyStruct(Protocol):
    Offset: int
    OffsetHigh: int


class _DummyUnion(Protocol):
    DUMMYSTRUCTNAME: _DummyStruct
    Pointer: object


class _Overlapped(Protocol):
    Internal: int
    InternalHigh: int
    DUMMYUNIONNAME: _DummyUnion
    hEvent: Handle


kernel32 = cast("_Kernel32", ffi.dlopen("kernel32.dll"))
ntdll = cast("_Nt", ffi.dlopen("ntdll.dll"))
ws2_32 = cast("_Ws2", ffi.dlopen("ws2_32.dll"))

################################################################
# Magic numbers
################################################################

# Here's a great resource for looking these up:
#   https://www.magnumdb.com
# (Tip: check the box to see "Hex value")

INVALID_HANDLE_VALUE = Handle(ffi.cast("HANDLE", -1))


class ErrorCodes(enum.IntEnum):
    STATUS_TIMEOUT = 0x102
    WAIT_TIMEOUT = 0x102
    WAIT_ABANDONED = 0x80
    WAIT_OBJECT_0 = 0x00  # object is signaled
    WAIT_FAILED = 0xFFFFFFFF
    ERROR_IO_PENDING = 997
    ERROR_OPERATION_ABORTED = 995
    ERROR_ABANDONED_WAIT_0 = 735
    ERROR_INVALID_HANDLE = 6
    ERROR_INVALID_PARAMETER = 87
    ERROR_NOT_FOUND = 1168
    ERROR_NOT_SOCKET = 10038


class FileFlags(enum.IntFlag):
    GENERIC_READ = 0x80000000
    SYNCHRONIZE = 0x00100000
    FILE_FLAG_OVERLAPPED = 0x40000000
    FILE_SHARE_READ = 1
    FILE_SHARE_WRITE = 2
    FILE_SHARE_DELETE = 4
    CREATE_NEW = 1
    CREATE_ALWAYS = 2
    OPEN_EXISTING = 3
    OPEN_ALWAYS = 4
    TRUNCATE_EXISTING = 5


class AFDPollFlags(enum.IntFlag):
    # These are drawn from a combination of:
    #   https://github.com/piscisaureus/wepoll/blob/master/src/afd.h
    #   https://github.com/reactos/reactos/blob/master/sdk/include/reactos/drivers/afd/shared.h
    AFD_POLL_RECEIVE = 0x0001
    AFD_POLL_RECEIVE_EXPEDITED = 0x0002  # OOB/urgent data
    AFD_POLL_SEND = 0x0004
    AFD_POLL_DISCONNECT = 0x0008  # received EOF (FIN)
    AFD_POLL_ABORT = 0x0010  # received RST
    AFD_POLL_LOCAL_CLOSE = 0x0020  # local socket object closed
    AFD_POLL_CONNECT = 0x0040  # socket is successfully connected
    AFD_POLL_ACCEPT = 0x0080  # you can call accept on this socket
    AFD_POLL_CONNECT_FAIL = 0x0100  # connect() terminated unsuccessfully
    # See WSAEventSelect docs for more details on these four:
    AFD_POLL_QOS = 0x0200
    AFD_POLL_GROUP_QOS = 0x0400
    AFD_POLL_ROUTING_INTERFACE_CHANGE = 0x0800
    AFD_POLL_EVENT_ADDRESS_LIST_CHANGE = 0x1000


class WSAIoctls(enum.IntEnum):
    SIO_BASE_HANDLE = 0x48000022
    SIO_BSP_HANDLE_SELECT = 0x4800001C
    SIO_BSP_HANDLE_POLL = 0x4800001D


class CompletionModes(enum.IntFlag):
    FILE_SKIP_COMPLETION_PORT_ON_SUCCESS = 0x1
    FILE_SKIP_SET_EVENT_ON_HANDLE = 0x2


class IoControlCodes(enum.IntEnum):
    IOCTL_AFD_POLL = 0x00012024


################################################################
# Generic helpers
################################################################


def _handle(obj: int | CData) -> Handle:
    # For now, represent handles as either cffi HANDLEs or as ints.  If you
    # try to pass in a file descriptor instead, it's not going to work
    # out. (For that msvcrt.get_osfhandle does the trick, but I don't know if
    # we'll actually need that for anything...) For sockets this doesn't
    # matter, Python never allocates an fd. So let's wait until we actually
    # encounter the problem before worrying about it.
    if isinstance(obj, int):
        return Handle(ffi.cast("HANDLE", obj))
    return Handle(obj)


def handle_array(count: int) -> HandleArray:
    """Make an array of handles."""
    return HandleArray(ffi.new(f"HANDLE[{count}]"))


def raise_winerror(
    winerror: int | None = None,
    *,
    filename: str | None = None,
    filename2: str | None = None,
) -> NoReturn:
    # assert sys.platform == "win32"  # TODO: make this work in MyPy
    # ... in the meanwhile, ffi.getwinerror() is undefined on non-Windows, necessitating the type
    # ignores.

    if winerror is None:
        err = ffi.getwinerror()  # type: ignore[attr-defined,unused-ignore]
        if err is None:
            raise RuntimeError("No error set?")
        winerror, msg = err
    else:
        err = ffi.getwinerror(winerror)  # type: ignore[attr-defined,unused-ignore]
        if err is None:
            raise RuntimeError("No error set?")
        _, msg = err
    # https://docs.python.org/3/library/exceptions.html#OSError
    raise OSError(0, msg, filename, winerror, filename2)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\urllib3\contrib\_securetransport\bindings.py ===
"""
This module uses ctypes to bind a whole bunch of functions and constants from
SecureTransport. The goal here is to provide the low-level API to
SecureTransport. These are essentially the C-level functions and constants, and
they're pretty gross to work with.

This code is a bastardised version of the code found in Will Bond's oscrypto
library. An enormous debt is owed to him for blazing this trail for us. For
that reason, this code should be considered to be covered both by urllib3's
license and by oscrypto's:

    Copyright (c) 2015-2016 Will Bond <will@wbond.net>

    Permission is hereby granted, free of charge, to any person obtaining a
    copy of this software and associated documentation files (the "Software"),
    to deal in the Software without restriction, including without limitation
    the rights to use, copy, modify, merge, publish, distribute, sublicense,
    and/or sell copies of the Software, and to permit persons to whom the
    Software is furnished to do so, subject to the following conditions:

    The above copyright notice and this permission notice shall be included in
    all copies or substantial portions of the Software.

    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
    AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
    LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
    FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
    DEALINGS IN THE SOFTWARE.
"""
from __future__ import absolute_import

import platform
from ctypes import (
    CDLL,
    CFUNCTYPE,
    POINTER,
    c_bool,
    c_byte,
    c_char_p,
    c_int32,
    c_long,
    c_size_t,
    c_uint32,
    c_ulong,
    c_void_p,
)
from ctypes.util import find_library

from ...packages.six import raise_from

if platform.system() != "Darwin":
    raise ImportError("Only macOS is supported")

version = platform.mac_ver()[0]
version_info = tuple(map(int, version.split(".")))
if version_info < (10, 8):
    raise OSError(
        "Only OS X 10.8 and newer are supported, not %s.%s"
        % (version_info[0], version_info[1])
    )


def load_cdll(name, macos10_16_path):
    """Loads a CDLL by name, falling back to known path on 10.16+"""
    try:
        # Big Sur is technically 11 but we use 10.16 due to the Big Sur
        # beta being labeled as 10.16.
        if version_info >= (10, 16):
            path = macos10_16_path
        else:
            path = find_library(name)
        if not path:
            raise OSError  # Caught and reraised as 'ImportError'
        return CDLL(path, use_errno=True)
    except OSError:
        raise_from(ImportError("The library %s failed to load" % name), None)


Security = load_cdll(
    "Security", "/System/Library/Frameworks/Security.framework/Security"
)
CoreFoundation = load_cdll(
    "CoreFoundation",
    "/System/Library/Frameworks/CoreFoundation.framework/CoreFoundation",
)


Boolean = c_bool
CFIndex = c_long
CFStringEncoding = c_uint32
CFData = c_void_p
CFString = c_void_p
CFArray = c_void_p
CFMutableArray = c_void_p
CFDictionary = c_void_p
CFError = c_void_p
CFType = c_void_p
CFTypeID = c_ulong

CFTypeRef = POINTER(CFType)
CFAllocatorRef = c_void_p

OSStatus = c_int32

CFDataRef = POINTER(CFData)
CFStringRef = POINTER(CFString)
CFArrayRef = POINTER(CFArray)
CFMutableArrayRef = POINTER(CFMutableArray)
CFDictionaryRef = POINTER(CFDictionary)
CFArrayCallBacks = c_void_p
CFDictionaryKeyCallBacks = c_void_p
CFDictionaryValueCallBacks = c_void_p

SecCertificateRef = POINTER(c_void_p)
SecExternalFormat = c_uint32
SecExternalItemType = c_uint32
SecIdentityRef = POINTER(c_void_p)
SecItemImportExportFlags = c_uint32
SecItemImportExportKeyParameters = c_void_p
SecKeychainRef = POINTER(c_void_p)
SSLProtocol = c_uint32
SSLCipherSuite = c_uint32
SSLContextRef = POINTER(c_void_p)
SecTrustRef = POINTER(c_void_p)
SSLConnectionRef = c_uint32
SecTrustResultType = c_uint32
SecTrustOptionFlags = c_uint32
SSLProtocolSide = c_uint32
SSLConnectionType = c_uint32
SSLSessionOption = c_uint32


try:
    Security.SecItemImport.argtypes = [
        CFDataRef,
        CFStringRef,
        POINTER(SecExternalFormat),
        POINTER(SecExternalItemType),
        SecItemImportExportFlags,
        POINTER(SecItemImportExportKeyParameters),
        SecKeychainRef,
        POINTER(CFArrayRef),
    ]
    Security.SecItemImport.restype = OSStatus

    Security.SecCertificateGetTypeID.argtypes = []
    Security.SecCertificateGetTypeID.restype = CFTypeID

    Security.SecIdentityGetTypeID.argtypes = []
    Security.SecIdentityGetTypeID.restype = CFTypeID

    Security.SecKeyGetTypeID.argtypes = []
    Security.SecKeyGetTypeID.restype = CFTypeID

    Security.SecCertificateCreateWithData.argtypes = [CFAllocatorRef, CFDataRef]
    Security.SecCertificateCreateWithData.restype = SecCertificateRef

    Security.SecCertificateCopyData.argtypes = [SecCertificateRef]
    Security.SecCertificateCopyData.restype = CFDataRef

    Security.SecCopyErrorMessageString.argtypes = [OSStatus, c_void_p]
    Security.SecCopyErrorMessageString.restype = CFStringRef

    Security.SecIdentityCreateWithCertificate.argtypes = [
        CFTypeRef,
        SecCertificateRef,
        POINTER(SecIdentityRef),
    ]
    Security.SecIdentityCreateWithCertificate.restype = OSStatus

    Security.SecKeychainCreate.argtypes = [
        c_char_p,
        c_uint32,
        c_void_p,
        Boolean,
        c_void_p,
        POINTER(SecKeychainRef),
    ]
    Security.SecKeychainCreate.restype = OSStatus

    Security.SecKeychainDelete.argtypes = [SecKeychainRef]
    Security.SecKeychainDelete.restype = OSStatus

    Security.SecPKCS12Import.argtypes = [
        CFDataRef,
        CFDictionaryRef,
        POINTER(CFArrayRef),
    ]
    Security.SecPKCS12Import.restype = OSStatus

    SSLReadFunc = CFUNCTYPE(OSStatus, SSLConnectionRef, c_void_p, POINTER(c_size_t))
    SSLWriteFunc = CFUNCTYPE(
        OSStatus, SSLConnectionRef, POINTER(c_byte), POINTER(c_size_t)
    )

    Security.SSLSetIOFuncs.argtypes = [SSLContextRef, SSLReadFunc, SSLWriteFunc]
    Security.SSLSetIOFuncs.restype = OSStatus

    Security.SSLSetPeerID.argtypes = [SSLContextRef, c_char_p, c_size_t]
    Security.SSLSetPeerID.restype = OSStatus

    Security.SSLSetCertificate.argtypes = [SSLContextRef, CFArrayRef]
    Security.SSLSetCertificate.restype = OSStatus

    Security.SSLSetCertificateAuthorities.argtypes = [SSLContextRef, CFTypeRef, Boolean]
    Security.SSLSetCertificateAuthorities.restype = OSStatus

    Security.SSLSetConnection.argtypes = [SSLContextRef, SSLConnectionRef]
    Security.SSLSetConnection.restype = OSStatus

    Security.SSLSetPeerDomainName.argtypes = [SSLContextRef, c_char_p, c_size_t]
    Security.SSLSetPeerDomainName.restype = OSStatus

    Security.SSLHandshake.argtypes = [SSLContextRef]
    Security.SSLHandshake.restype = OSStatus

    Security.SSLRead.argtypes = [SSLContextRef, c_char_p, c_size_t, POINTER(c_size_t)]
    Security.SSLRead.restype = OSStatus

    Security.SSLWrite.argtypes = [SSLContextRef, c_char_p, c_size_t, POINTER(c_size_t)]
    Security.SSLWrite.restype = OSStatus

    Security.SSLClose.argtypes = [SSLContextRef]
    Security.SSLClose.restype = OSStatus

    Security.SSLGetNumberSupportedCiphers.argtypes = [SSLContextRef, POINTER(c_size_t)]
    Security.SSLGetNumberSupportedCiphers.restype = OSStatus

    Security.SSLGetSupportedCiphers.argtypes = [
        SSLContextRef,
        POINTER(SSLCipherSuite),
        POINTER(c_size_t),
    ]
    Security.SSLGetSupportedCiphers.restype = OSStatus

    Security.SSLSetEnabledCiphers.argtypes = [
        SSLContextRef,
        POINTER(SSLCipherSuite),
        c_size_t,
    ]
    Security.SSLSetEnabledCiphers.restype = OSStatus

    Security.SSLGetNumberEnabledCiphers.argtype = [SSLContextRef, POINTER(c_size_t)]
    Security.SSLGetNumberEnabledCiphers.restype = OSStatus

    Security.SSLGetEnabledCiphers.argtypes = [
        SSLContextRef,
        POINTER(SSLCipherSuite),
        POINTER(c_size_t),
    ]
    Security.SSLGetEnabledCiphers.restype = OSStatus

    Security.SSLGetNegotiatedCipher.argtypes = [SSLContextRef, POINTER(SSLCipherSuite)]
    Security.SSLGetNegotiatedCipher.restype = OSStatus

    Security.SSLGetNegotiatedProtocolVersion.argtypes = [
        SSLContextRef,
        POINTER(SSLProtocol),
    ]
    Security.SSLGetNegotiatedProtocolVersion.restype = OSStatus

    Security.SSLCopyPeerTrust.argtypes = [SSLContextRef, POINTER(SecTrustRef)]
    Security.SSLCopyPeerTrust.restype = OSStatus

    Security.SecTrustSetAnchorCertificates.argtypes = [SecTrustRef, CFArrayRef]
    Security.SecTrustSetAnchorCertificates.restype = OSStatus

    Security.SecTrustSetAnchorCertificatesOnly.argstypes = [SecTrustRef, Boolean]
    Security.SecTrustSetAnchorCertificatesOnly.restype = OSStatus

    Security.SecTrustEvaluate.argtypes = [SecTrustRef, POINTER(SecTrustResultType)]
    Security.SecTrustEvaluate.restype = OSStatus

    Security.SecTrustGetCertificateCount.argtypes = [SecTrustRef]
    Security.SecTrustGetCertificateCount.restype = CFIndex

    Security.SecTrustGetCertificateAtIndex.argtypes = [SecTrustRef, CFIndex]
    Security.SecTrustGetCertificateAtIndex.restype = SecCertificateRef

    Security.SSLCreateContext.argtypes = [
        CFAllocatorRef,
        SSLProtocolSide,
        SSLConnectionType,
    ]
    Security.SSLCreateContext.restype = SSLContextRef

    Security.SSLSetSessionOption.argtypes = [SSLContextRef, SSLSessionOption, Boolean]
    Security.SSLSetSessionOption.restype = OSStatus

    Security.SSLSetProtocolVersionMin.argtypes = [SSLContextRef, SSLProtocol]
    Security.SSLSetProtocolVersionMin.restype = OSStatus

    Security.SSLSetProtocolVersionMax.argtypes = [SSLContextRef, SSLProtocol]
    Security.SSLSetProtocolVersionMax.restype = OSStatus

    try:
        Security.SSLSetALPNProtocols.argtypes = [SSLContextRef, CFArrayRef]
        Security.SSLSetALPNProtocols.restype = OSStatus
    except AttributeError:
        # Supported only in 10.12+
        pass

    Security.SecCopyErrorMessageString.argtypes = [OSStatus, c_void_p]
    Security.SecCopyErrorMessageString.restype = CFStringRef

    Security.SSLReadFunc = SSLReadFunc
    Security.SSLWriteFunc = SSLWriteFunc
    Security.SSLContextRef = SSLContextRef
    Security.SSLProtocol = SSLProtocol
    Security.SSLCipherSuite = SSLCipherSuite
    Security.SecIdentityRef = SecIdentityRef
    Security.SecKeychainRef = SecKeychainRef
    Security.SecTrustRef = SecTrustRef
    Security.SecTrustResultType = SecTrustResultType
    Security.SecExternalFormat = SecExternalFormat
    Security.OSStatus = OSStatus

    Security.kSecImportExportPassphrase = CFStringRef.in_dll(
        Security, "kSecImportExportPassphrase"
    )
    Security.kSecImportItemIdentity = CFStringRef.in_dll(
        Security, "kSecImportItemIdentity"
    )

    # CoreFoundation time!
    CoreFoundation.CFRetain.argtypes = [CFTypeRef]
    CoreFoundation.CFRetain.restype = CFTypeRef

    CoreFoundation.CFRelease.argtypes = [CFTypeRef]
    CoreFoundation.CFRelease.restype = None

    CoreFoundation.CFGetTypeID.argtypes = [CFTypeRef]
    CoreFoundation.CFGetTypeID.restype = CFTypeID

    CoreFoundation.CFStringCreateWithCString.argtypes = [
        CFAllocatorRef,
        c_char_p,
        CFStringEncoding,
    ]
    CoreFoundation.CFStringCreateWithCString.restype = CFStringRef

    CoreFoundation.CFStringGetCStringPtr.argtypes = [CFStringRef, CFStringEncoding]
    CoreFoundation.CFStringGetCStringPtr.restype = c_char_p

    CoreFoundation.CFStringGetCString.argtypes = [
        CFStringRef,
        c_char_p,
        CFIndex,
        CFStringEncoding,
    ]
    CoreFoundation.CFStringGetCString.restype = c_bool

    CoreFoundation.CFDataCreate.argtypes = [CFAllocatorRef, c_char_p, CFIndex]
    CoreFoundation.CFDataCreate.restype = CFDataRef

    CoreFoundation.CFDataGetLength.argtypes = [CFDataRef]
    CoreFoundation.CFDataGetLength.restype = CFIndex

    CoreFoundation.CFDataGetBytePtr.argtypes = [CFDataRef]
    CoreFoundation.CFDataGetBytePtr.restype = c_void_p

    CoreFoundation.CFDictionaryCreate.argtypes = [
        CFAllocatorRef,
        POINTER(CFTypeRef),
        POINTER(CFTypeRef),
        CFIndex,
        CFDictionaryKeyCallBacks,
        CFDictionaryValueCallBacks,
    ]
    CoreFoundation.CFDictionaryCreate.restype = CFDictionaryRef

    CoreFoundation.CFDictionaryGetValue.argtypes = [CFDictionaryRef, CFTypeRef]
    CoreFoundation.CFDictionaryGetValue.restype = CFTypeRef

    CoreFoundation.CFArrayCreate.argtypes = [
        CFAllocatorRef,
        POINTER(CFTypeRef),
        CFIndex,
        CFArrayCallBacks,
    ]
    CoreFoundation.CFArrayCreate.restype = CFArrayRef

    CoreFoundation.CFArrayCreateMutable.argtypes = [
        CFAllocatorRef,
        CFIndex,
        CFArrayCallBacks,
    ]
    CoreFoundation.CFArrayCreateMutable.restype = CFMutableArrayRef

    CoreFoundation.CFArrayAppendValue.argtypes = [CFMutableArrayRef, c_void_p]
    CoreFoundation.CFArrayAppendValue.restype = None

    CoreFoundation.CFArrayGetCount.argtypes = [CFArrayRef]
    CoreFoundation.CFArrayGetCount.restype = CFIndex

    CoreFoundation.CFArrayGetValueAtIndex.argtypes = [CFArrayRef, CFIndex]
    CoreFoundation.CFArrayGetValueAtIndex.restype = c_void_p

    CoreFoundation.kCFAllocatorDefault = CFAllocatorRef.in_dll(
        CoreFoundation, "kCFAllocatorDefault"
    )
    CoreFoundation.kCFTypeArrayCallBacks = c_void_p.in_dll(
        CoreFoundation, "kCFTypeArrayCallBacks"
    )
    CoreFoundation.kCFTypeDictionaryKeyCallBacks = c_void_p.in_dll(
        CoreFoundation, "kCFTypeDictionaryKeyCallBacks"
    )
    CoreFoundation.kCFTypeDictionaryValueCallBacks = c_void_p.in_dll(
        CoreFoundation, "kCFTypeDictionaryValueCallBacks"
    )

    CoreFoundation.CFTypeRef = CFTypeRef
    CoreFoundation.CFArrayRef = CFArrayRef
    CoreFoundation.CFStringRef = CFStringRef
    CoreFoundation.CFDictionaryRef = CFDictionaryRef

except (AttributeError):
    raise ImportError("Error initializing ctypes")


class CFConst(object):
    """
    A class object that acts as essentially a namespace for CoreFoundation
    constants.
    """

    kCFStringEncodingUTF8 = CFStringEncoding(0x08000100)


class SecurityConst(object):
    """
    A class object that acts as essentially a namespace for Security constants.
    """

    kSSLSessionOptionBreakOnServerAuth = 0

    kSSLProtocol2 = 1
    kSSLProtocol3 = 2
    kTLSProtocol1 = 4
    kTLSProtocol11 = 7
    kTLSProtocol12 = 8
    # SecureTransport does not support TLS 1.3 even if there's a constant for it
    kTLSProtocol13 = 10
    kTLSProtocolMaxSupported = 999

    kSSLClientSide = 1
    kSSLStreamType = 0

    kSecFormatPEMSequence = 10

    kSecTrustResultInvalid = 0
    kSecTrustResultProceed = 1
    # This gap is present on purpose: this was kSecTrustResultConfirm, which
    # is deprecated.
    kSecTrustResultDeny = 3
    kSecTrustResultUnspecified = 4
    kSecTrustResultRecoverableTrustFailure = 5
    kSecTrustResultFatalTrustFailure = 6
    kSecTrustResultOtherError = 7

    errSSLProtocol = -9800
    errSSLWouldBlock = -9803
    errSSLClosedGraceful = -9805
    errSSLClosedNoNotify = -9816
    errSSLClosedAbort = -9806

    errSSLXCertChainInvalid = -9807
    errSSLCrypto = -9809
    errSSLInternal = -9810
    errSSLCertExpired = -9814
    errSSLCertNotYetValid = -9815
    errSSLUnknownRootCert = -9812
    errSSLNoRootCert = -9813
    errSSLHostNameMismatch = -9843
    errSSLPeerHandshakeFail = -9824
    errSSLPeerUserCancelled = -9839
    errSSLWeakPeerEphemeralDHKey = -9850
    errSSLServerAuthCompleted = -9841
    errSSLRecordOverflow = -9847

    errSecVerifyFailed = -67808
    errSecNoTrustSettings = -25263
    errSecItemNotFound = -25300
    errSecInvalidTrustSettings = -25262

    # Cipher suites. We only pick the ones our default cipher string allows.
    # Source: https://developer.apple.com/documentation/security/1550981-ssl_cipher_suite_values
    TLS_ECDHE_ECDSA_WITH_AES_256_GCM_SHA384 = 0xC02C
    TLS_ECDHE_RSA_WITH_AES_256_GCM_SHA384 = 0xC030
    TLS_ECDHE_ECDSA_WITH_AES_128_GCM_SHA256 = 0xC02B
    TLS_ECDHE_RSA_WITH_AES_128_GCM_SHA256 = 0xC02F
    TLS_ECDHE_ECDSA_WITH_CHACHA20_POLY1305_SHA256 = 0xCCA9
    TLS_ECDHE_RSA_WITH_CHACHA20_POLY1305_SHA256 = 0xCCA8
    TLS_DHE_RSA_WITH_AES_256_GCM_SHA384 = 0x009F
    TLS_DHE_RSA_WITH_AES_128_GCM_SHA256 = 0x009E
    TLS_ECDHE_ECDSA_WITH_AES_256_CBC_SHA384 = 0xC024
    TLS_ECDHE_RSA_WITH_AES_256_CBC_SHA384 = 0xC028
    TLS_ECDHE_ECDSA_WITH_AES_256_CBC_SHA = 0xC00A
    TLS_ECDHE_RSA_WITH_AES_256_CBC_SHA = 0xC014
    TLS_DHE_RSA_WITH_AES_256_CBC_SHA256 = 0x006B
    TLS_DHE_RSA_WITH_AES_256_CBC_SHA = 0x0039
    TLS_ECDHE_ECDSA_WITH_AES_128_CBC_SHA256 = 0xC023
    TLS_ECDHE_RSA_WITH_AES_128_CBC_SHA256 = 0xC027
    TLS_ECDHE_ECDSA_WITH_AES_128_CBC_SHA = 0xC009
    TLS_ECDHE_RSA_WITH_AES_128_CBC_SHA = 0xC013
    TLS_DHE_RSA_WITH_AES_128_CBC_SHA256 = 0x0067
    TLS_DHE_RSA_WITH_AES_128_CBC_SHA = 0x0033
    TLS_RSA_WITH_AES_256_GCM_SHA384 = 0x009D
    TLS_RSA_WITH_AES_128_GCM_SHA256 = 0x009C
    TLS_RSA_WITH_AES_256_CBC_SHA256 = 0x003D
    TLS_RSA_WITH_AES_128_CBC_SHA256 = 0x003C
    TLS_RSA_WITH_AES_256_CBC_SHA = 0x0035
    TLS_RSA_WITH_AES_128_CBC_SHA = 0x002F
    TLS_AES_128_GCM_SHA256 = 0x1301
    TLS_AES_256_GCM_SHA384 = 0x1302
    TLS_AES_128_CCM_8_SHA256 = 0x1305
    TLS_AES_128_CCM_SHA256 = 0x1304

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\stats\rv_interface.py ===
from sympy.sets import FiniteSet
from sympy.core.numbers import Rational
from sympy.core.relational import Eq
from sympy.core.symbol import Dummy
from sympy.functions.combinatorial.factorials import FallingFactorial
from sympy.functions.elementary.exponential import (exp, log)
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.piecewise import piecewise_fold
from sympy.integrals.integrals import Integral
from sympy.solvers.solveset import solveset
from .rv import (probability, expectation, density, where, given, pspace, cdf, PSpace,
                 characteristic_function, sample, sample_iter, random_symbols, independent, dependent,
                 sampling_density, moment_generating_function, quantile, is_random,
                 sample_stochastic_process)


__all__ = ['P', 'E', 'H', 'density', 'where', 'given', 'sample', 'cdf',
        'characteristic_function', 'pspace', 'sample_iter', 'variance', 'std',
        'skewness', 'kurtosis', 'covariance', 'dependent', 'entropy', 'median',
        'independent', 'random_symbols', 'correlation', 'factorial_moment',
        'moment', 'cmoment', 'sampling_density', 'moment_generating_function',
        'smoment', 'quantile', 'sample_stochastic_process']



def moment(X, n, c=0, condition=None, *, evaluate=True, **kwargs):
    """
    Return the nth moment of a random expression about c.

    .. math::
        moment(X, c, n) = E((X-c)^{n})

    Default value of c is 0.

    Examples
    ========

    >>> from sympy.stats import Die, moment, E
    >>> X = Die('X', 6)
    >>> moment(X, 1, 6)
    -5/2
    >>> moment(X, 2)
    91/6
    >>> moment(X, 1) == E(X)
    True
    """
    from sympy.stats.symbolic_probability import Moment
    if evaluate:
        return Moment(X, n, c, condition).doit()
    return Moment(X, n, c, condition).rewrite(Integral)


def variance(X, condition=None, **kwargs):
    """
    Variance of a random expression.

    .. math::
        variance(X) = E((X-E(X))^{2})

    Examples
    ========

    >>> from sympy.stats import Die, Bernoulli, variance
    >>> from sympy import simplify, Symbol

    >>> X = Die('X', 6)
    >>> p = Symbol('p')
    >>> B = Bernoulli('B', p, 1, 0)

    >>> variance(2*X)
    35/3

    >>> simplify(variance(B))
    p*(1 - p)
    """
    if is_random(X) and pspace(X) == PSpace():
        from sympy.stats.symbolic_probability import Variance
        return Variance(X, condition)

    return cmoment(X, 2, condition, **kwargs)


def standard_deviation(X, condition=None, **kwargs):
    r"""
    Standard Deviation of a random expression

    .. math::
        std(X) = \sqrt(E((X-E(X))^{2}))

    Examples
    ========

    >>> from sympy.stats import Bernoulli, std
    >>> from sympy import Symbol, simplify

    >>> p = Symbol('p')
    >>> B = Bernoulli('B', p, 1, 0)

    >>> simplify(std(B))
    sqrt(p*(1 - p))
    """
    return sqrt(variance(X, condition, **kwargs))
std = standard_deviation

def entropy(expr, condition=None, **kwargs):
    """
    Calculates entropy of a probability distribution.

    Parameters
    ==========

    expression : the random expression whose entropy is to be calculated
    condition : optional, to specify conditions on random expression
    b: base of the logarithm, optional
       By default, it is taken as Euler's number

    Returns
    =======

    result : Entropy of the expression, a constant

    Examples
    ========

    >>> from sympy.stats import Normal, Die, entropy
    >>> X = Normal('X', 0, 1)
    >>> entropy(X)
    log(2)/2 + 1/2 + log(pi)/2

    >>> D = Die('D', 4)
    >>> entropy(D)
    log(4)

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Entropy_%28information_theory%29
    .. [2] https://www.crmarsh.com/static/pdf/Charles_Marsh_Continuous_Entropy.pdf
    .. [3] https://kconrad.math.uconn.edu/blurbs/analysis/entropypost.pdf
    """
    pdf = density(expr, condition, **kwargs)
    base = kwargs.get('b', exp(1))
    if isinstance(pdf, dict):
            return sum(-prob*log(prob, base) for prob in pdf.values())
    return expectation(-log(pdf(expr), base))

def covariance(X, Y, condition=None, **kwargs):
    """
    Covariance of two random expressions.

    Explanation
    ===========

    The expectation that the two variables will rise and fall together

    .. math::
        covariance(X,Y) = E((X-E(X)) (Y-E(Y)))

    Examples
    ========

    >>> from sympy.stats import Exponential, covariance
    >>> from sympy import Symbol

    >>> rate = Symbol('lambda', positive=True, real=True)
    >>> X = Exponential('X', rate)
    >>> Y = Exponential('Y', rate)

    >>> covariance(X, X)
    lambda**(-2)
    >>> covariance(X, Y)
    0
    >>> covariance(X, Y + rate*X)
    1/lambda
    """
    if (is_random(X) and pspace(X) == PSpace()) or (is_random(Y) and pspace(Y) == PSpace()):
        from sympy.stats.symbolic_probability import Covariance
        return Covariance(X, Y, condition)

    return expectation(
        (X - expectation(X, condition, **kwargs)) *
        (Y - expectation(Y, condition, **kwargs)),
        condition, **kwargs)


def correlation(X, Y, condition=None, **kwargs):
    r"""
    Correlation of two random expressions, also known as correlation
    coefficient or Pearson's correlation.

    Explanation
    ===========

    The normalized expectation that the two variables will rise
    and fall together

    .. math::
        correlation(X,Y) = E((X-E(X))(Y-E(Y)) / (\sigma_x  \sigma_y))

    Examples
    ========

    >>> from sympy.stats import Exponential, correlation
    >>> from sympy import Symbol

    >>> rate = Symbol('lambda', positive=True, real=True)
    >>> X = Exponential('X', rate)
    >>> Y = Exponential('Y', rate)

    >>> correlation(X, X)
    1
    >>> correlation(X, Y)
    0
    >>> correlation(X, Y + rate*X)
    1/sqrt(1 + lambda**(-2))
    """
    return covariance(X, Y, condition, **kwargs)/(std(X, condition, **kwargs)
     * std(Y, condition, **kwargs))


def cmoment(X, n, condition=None, *, evaluate=True, **kwargs):
    """
    Return the nth central moment of a random expression about its mean.

    .. math::
        cmoment(X, n) = E((X - E(X))^{n})

    Examples
    ========

    >>> from sympy.stats import Die, cmoment, variance
    >>> X = Die('X', 6)
    >>> cmoment(X, 3)
    0
    >>> cmoment(X, 2)
    35/12
    >>> cmoment(X, 2) == variance(X)
    True
    """
    from sympy.stats.symbolic_probability import CentralMoment
    if evaluate:
        return CentralMoment(X, n, condition).doit()
    return CentralMoment(X, n, condition).rewrite(Integral)


def smoment(X, n, condition=None, **kwargs):
    r"""
    Return the nth Standardized moment of a random expression.

    .. math::
        smoment(X, n) = E(((X - \mu)/\sigma_X)^{n})

    Examples
    ========

    >>> from sympy.stats import skewness, Exponential, smoment
    >>> from sympy import Symbol
    >>> rate = Symbol('lambda', positive=True, real=True)
    >>> Y = Exponential('Y', rate)
    >>> smoment(Y, 4)
    9
    >>> smoment(Y, 4) == smoment(3*Y, 4)
    True
    >>> smoment(Y, 3) == skewness(Y)
    True
    """
    sigma = std(X, condition, **kwargs)
    return (1/sigma)**n*cmoment(X, n, condition, **kwargs)

def skewness(X, condition=None, **kwargs):
    r"""
    Measure of the asymmetry of the probability distribution.

    Explanation
    ===========

    Positive skew indicates that most of the values lie to the right of
    the mean.

    .. math::
        skewness(X) = E(((X - E(X))/\sigma_X)^{3})

    Parameters
    ==========

    condition : Expr containing RandomSymbols
            A conditional expression. skewness(X, X>0) is skewness of X given X > 0

    Examples
    ========

    >>> from sympy.stats import skewness, Exponential, Normal
    >>> from sympy import Symbol
    >>> X = Normal('X', 0, 1)
    >>> skewness(X)
    0
    >>> skewness(X, X > 0) # find skewness given X > 0
    (-sqrt(2)/sqrt(pi) + 4*sqrt(2)/pi**(3/2))/(1 - 2/pi)**(3/2)

    >>> rate = Symbol('lambda', positive=True, real=True)
    >>> Y = Exponential('Y', rate)
    >>> skewness(Y)
    2
    """
    return smoment(X, 3, condition=condition, **kwargs)

def kurtosis(X, condition=None, **kwargs):
    r"""
    Characterizes the tails/outliers of a probability distribution.

    Explanation
    ===========

    Kurtosis of any univariate normal distribution is 3. Kurtosis less than
    3 means that the distribution produces fewer and less extreme outliers
    than the normal distribution.

    .. math::
        kurtosis(X) = E(((X - E(X))/\sigma_X)^{4})

    Parameters
    ==========

    condition : Expr containing RandomSymbols
            A conditional expression. kurtosis(X, X>0) is kurtosis of X given X > 0

    Examples
    ========

    >>> from sympy.stats import kurtosis, Exponential, Normal
    >>> from sympy import Symbol
    >>> X = Normal('X', 0, 1)
    >>> kurtosis(X)
    3
    >>> kurtosis(X, X > 0) # find kurtosis given X > 0
    (-4/pi - 12/pi**2 + 3)/(1 - 2/pi)**2

    >>> rate = Symbol('lamda', positive=True, real=True)
    >>> Y = Exponential('Y', rate)
    >>> kurtosis(Y)
    9

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Kurtosis
    .. [2] https://mathworld.wolfram.com/Kurtosis.html
    """
    return smoment(X, 4, condition=condition, **kwargs)


def factorial_moment(X, n, condition=None, **kwargs):
    """
    The factorial moment is a mathematical quantity defined as the expectation
    or average of the falling factorial of a random variable.

    .. math::
        factorial-moment(X, n) = E(X(X - 1)(X - 2)...(X - n + 1))

    Parameters
    ==========

    n: A natural number, n-th factorial moment.

    condition : Expr containing RandomSymbols
            A conditional expression.

    Examples
    ========

    >>> from sympy.stats import factorial_moment, Poisson, Binomial
    >>> from sympy import Symbol, S
    >>> lamda = Symbol('lamda')
    >>> X = Poisson('X', lamda)
    >>> factorial_moment(X, 2)
    lamda**2
    >>> Y = Binomial('Y', 2, S.Half)
    >>> factorial_moment(Y, 2)
    1/2
    >>> factorial_moment(Y, 2, Y > 1) # find factorial moment for Y > 1
    2

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Factorial_moment
    .. [2] https://mathworld.wolfram.com/FactorialMoment.html
    """
    return expectation(FallingFactorial(X, n), condition=condition, **kwargs)

def median(X, evaluate=True, **kwargs):
    r"""
    Calculates the median of the probability distribution.

    Explanation
    ===========

    Mathematically, median of Probability distribution is defined as all those
    values of `m` for which the following condition is satisfied

    .. math::
        P(X\leq m) \geq  \frac{1}{2} \text{ and} \text{ } P(X\geq m)\geq \frac{1}{2}

    Parameters
    ==========

    X: The random expression whose median is to be calculated.

    Returns
    =======

    The FiniteSet or an Interval which contains the median of the
    random expression.

    Examples
    ========

    >>> from sympy.stats import Normal, Die, median
    >>> N = Normal('N', 3, 1)
    >>> median(N)
    {3}
    >>> D = Die('D')
    >>> median(D)
    {3, 4}

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Median#Probability_distributions

    """
    if not is_random(X):
        return X

    from sympy.stats.crv import ContinuousPSpace
    from sympy.stats.drv import DiscretePSpace
    from sympy.stats.frv import FinitePSpace

    if isinstance(pspace(X), FinitePSpace):
        cdf = pspace(X).compute_cdf(X)
        result = []
        for key, value in cdf.items():
            if value>= Rational(1, 2) and (1 - value) + \
            pspace(X).probability(Eq(X, key)) >= Rational(1, 2):
                result.append(key)
        return FiniteSet(*result)
    if isinstance(pspace(X), (ContinuousPSpace, DiscretePSpace)):
        cdf = pspace(X).compute_cdf(X)
        x = Dummy('x')
        result = solveset(piecewise_fold(cdf(x) - Rational(1, 2)), x, pspace(X).set)
        return result
    raise NotImplementedError("The median of %s is not implemented."%str(pspace(X)))


def coskewness(X, Y, Z, condition=None, **kwargs):
    r"""
    Calculates the co-skewness of three random variables.

    Explanation
    ===========

    Mathematically Coskewness is defined as

    .. math::
        coskewness(X,Y,Z)=\frac{E[(X-E[X]) * (Y-E[Y]) * (Z-E[Z])]} {\sigma_{X}\sigma_{Y}\sigma_{Z}}

    Parameters
    ==========

    X : RandomSymbol
            Random Variable used to calculate coskewness
    Y : RandomSymbol
            Random Variable used to calculate coskewness
    Z : RandomSymbol
            Random Variable used to calculate coskewness
    condition : Expr containing RandomSymbols
            A conditional expression

    Examples
    ========

    >>> from sympy.stats import coskewness, Exponential, skewness
    >>> from sympy import symbols
    >>> p = symbols('p', positive=True)
    >>> X = Exponential('X', p)
    >>> Y = Exponential('Y', 2*p)
    >>> coskewness(X, Y, Y)
    0
    >>> coskewness(X, Y + X, Y + 2*X)
    16*sqrt(85)/85
    >>> coskewness(X + 2*Y, Y + X, Y + 2*X, X > 3)
    9*sqrt(170)/85
    >>> coskewness(Y, Y, Y) == skewness(Y)
    True
    >>> coskewness(X, Y + p*X, Y + 2*p*X)
    4/(sqrt(1 + 1/(4*p**2))*sqrt(4 + 1/(4*p**2)))

    Returns
    =======

    coskewness : The coskewness of the three random variables

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Coskewness

    """
    num = expectation((X - expectation(X, condition, **kwargs)) \
         * (Y - expectation(Y, condition, **kwargs)) \
         * (Z - expectation(Z, condition, **kwargs)), condition, **kwargs)
    den = std(X, condition, **kwargs) * std(Y, condition, **kwargs) \
         * std(Z, condition, **kwargs)
    return num/den


P = probability
E = expectation
H = entropy

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\urllib3\contrib\pyopenssl.py ===
"""
TLS with SNI_-support for Python 2. Follow these instructions if you would
like to verify TLS certificates in Python 2. Note, the default libraries do
*not* do certificate checking; you need to do additional work to validate
certificates yourself.

This needs the following packages installed:

* `pyOpenSSL`_ (tested with 16.0.0)
* `cryptography`_ (minimum 1.3.4, from pyopenssl)
* `idna`_ (minimum 2.0, from cryptography)

However, pyopenssl depends on cryptography, which depends on idna, so while we
use all three directly here we end up having relatively few packages required.

You can install them with the following command:

.. code-block:: bash

    $ python -m pip install pyopenssl cryptography idna

To activate certificate checking, call
:func:`~urllib3.contrib.pyopenssl.inject_into_urllib3` from your Python code
before you begin making HTTP requests. This can be done in a ``sitecustomize``
module, or at any other time before your application begins using ``urllib3``,
like this:

.. code-block:: python

    try:
        import pip._vendor.urllib3.contrib.pyopenssl as pyopenssl
        pyopenssl.inject_into_urllib3()
    except ImportError:
        pass

Now you can use :mod:`urllib3` as you normally would, and it will support SNI
when the required modules are installed.

Activating this module also has the positive side effect of disabling SSL/TLS
compression in Python 2 (see `CRIME attack`_).

.. _sni: https://en.wikipedia.org/wiki/Server_Name_Indication
.. _crime attack: https://en.wikipedia.org/wiki/CRIME_(security_exploit)
.. _pyopenssl: https://www.pyopenssl.org
.. _cryptography: https://cryptography.io
.. _idna: https://github.com/kjd/idna
"""
from __future__ import absolute_import

import OpenSSL.crypto
import OpenSSL.SSL
from cryptography import x509
from cryptography.hazmat.backends.openssl import backend as openssl_backend

try:
    from cryptography.x509 import UnsupportedExtension
except ImportError:
    # UnsupportedExtension is gone in cryptography >= 2.1.0
    class UnsupportedExtension(Exception):
        pass


from io import BytesIO
from socket import error as SocketError
from socket import timeout

try:  # Platform-specific: Python 2
    from socket import _fileobject
except ImportError:  # Platform-specific: Python 3
    _fileobject = None
    from ..packages.backports.makefile import backport_makefile

import logging
import ssl
import sys
import warnings

from .. import util
from ..packages import six
from ..util.ssl_ import PROTOCOL_TLS_CLIENT

warnings.warn(
    "'urllib3.contrib.pyopenssl' module is deprecated and will be removed "
    "in a future release of urllib3 2.x. Read more in this issue: "
    "https://github.com/urllib3/urllib3/issues/2680",
    category=DeprecationWarning,
    stacklevel=2,
)

__all__ = ["inject_into_urllib3", "extract_from_urllib3"]

# SNI always works.
HAS_SNI = True

# Map from urllib3 to PyOpenSSL compatible parameter-values.
_openssl_versions = {
    util.PROTOCOL_TLS: OpenSSL.SSL.SSLv23_METHOD,
    PROTOCOL_TLS_CLIENT: OpenSSL.SSL.SSLv23_METHOD,
    ssl.PROTOCOL_TLSv1: OpenSSL.SSL.TLSv1_METHOD,
}

if hasattr(ssl, "PROTOCOL_SSLv3") and hasattr(OpenSSL.SSL, "SSLv3_METHOD"):
    _openssl_versions[ssl.PROTOCOL_SSLv3] = OpenSSL.SSL.SSLv3_METHOD

if hasattr(ssl, "PROTOCOL_TLSv1_1") and hasattr(OpenSSL.SSL, "TLSv1_1_METHOD"):
    _openssl_versions[ssl.PROTOCOL_TLSv1_1] = OpenSSL.SSL.TLSv1_1_METHOD

if hasattr(ssl, "PROTOCOL_TLSv1_2") and hasattr(OpenSSL.SSL, "TLSv1_2_METHOD"):
    _openssl_versions[ssl.PROTOCOL_TLSv1_2] = OpenSSL.SSL.TLSv1_2_METHOD


_stdlib_to_openssl_verify = {
    ssl.CERT_NONE: OpenSSL.SSL.VERIFY_NONE,
    ssl.CERT_OPTIONAL: OpenSSL.SSL.VERIFY_PEER,
    ssl.CERT_REQUIRED: OpenSSL.SSL.VERIFY_PEER
    + OpenSSL.SSL.VERIFY_FAIL_IF_NO_PEER_CERT,
}
_openssl_to_stdlib_verify = dict((v, k) for k, v in _stdlib_to_openssl_verify.items())

# OpenSSL will only write 16K at a time
SSL_WRITE_BLOCKSIZE = 16384

orig_util_HAS_SNI = util.HAS_SNI
orig_util_SSLContext = util.ssl_.SSLContext


log = logging.getLogger(__name__)


def inject_into_urllib3():
    "Monkey-patch urllib3 with PyOpenSSL-backed SSL-support."

    _validate_dependencies_met()

    util.SSLContext = PyOpenSSLContext
    util.ssl_.SSLContext = PyOpenSSLContext
    util.HAS_SNI = HAS_SNI
    util.ssl_.HAS_SNI = HAS_SNI
    util.IS_PYOPENSSL = True
    util.ssl_.IS_PYOPENSSL = True


def extract_from_urllib3():
    "Undo monkey-patching by :func:`inject_into_urllib3`."

    util.SSLContext = orig_util_SSLContext
    util.ssl_.SSLContext = orig_util_SSLContext
    util.HAS_SNI = orig_util_HAS_SNI
    util.ssl_.HAS_SNI = orig_util_HAS_SNI
    util.IS_PYOPENSSL = False
    util.ssl_.IS_PYOPENSSL = False


def _validate_dependencies_met():
    """
    Verifies that PyOpenSSL's package-level dependencies have been met.
    Throws `ImportError` if they are not met.
    """
    # Method added in `cryptography==1.1`; not available in older versions
    from cryptography.x509.extensions import Extensions

    if getattr(Extensions, "get_extension_for_class", None) is None:
        raise ImportError(
            "'cryptography' module missing required functionality.  "
            "Try upgrading to v1.3.4 or newer."
        )

    # pyOpenSSL 0.14 and above use cryptography for OpenSSL bindings. The _x509
    # attribute is only present on those versions.
    from OpenSSL.crypto import X509

    x509 = X509()
    if getattr(x509, "_x509", None) is None:
        raise ImportError(
            "'pyOpenSSL' module missing required functionality. "
            "Try upgrading to v0.14 or newer."
        )


def _dnsname_to_stdlib(name):
    """
    Converts a dNSName SubjectAlternativeName field to the form used by the
    standard library on the given Python version.

    Cryptography produces a dNSName as a unicode string that was idna-decoded
    from ASCII bytes. We need to idna-encode that string to get it back, and
    then on Python 3 we also need to convert to unicode via UTF-8 (the stdlib
    uses PyUnicode_FromStringAndSize on it, which decodes via UTF-8).

    If the name cannot be idna-encoded then we return None signalling that
    the name given should be skipped.
    """

    def idna_encode(name):
        """
        Borrowed wholesale from the Python Cryptography Project. It turns out
        that we can't just safely call `idna.encode`: it can explode for
        wildcard names. This avoids that problem.
        """
        from pip._vendor import idna

        try:
            for prefix in [u"*.", u"."]:
                if name.startswith(prefix):
                    name = name[len(prefix) :]
                    return prefix.encode("ascii") + idna.encode(name)
            return idna.encode(name)
        except idna.core.IDNAError:
            return None

    # Don't send IPv6 addresses through the IDNA encoder.
    if ":" in name:
        return name

    name = idna_encode(name)
    if name is None:
        return None
    elif sys.version_info >= (3, 0):
        name = name.decode("utf-8")
    return name


def get_subj_alt_name(peer_cert):
    """
    Given an PyOpenSSL certificate, provides all the subject alternative names.
    """
    # Pass the cert to cryptography, which has much better APIs for this.
    if hasattr(peer_cert, "to_cryptography"):
        cert = peer_cert.to_cryptography()
    else:
        der = OpenSSL.crypto.dump_certificate(OpenSSL.crypto.FILETYPE_ASN1, peer_cert)
        cert = x509.load_der_x509_certificate(der, openssl_backend)

    # We want to find the SAN extension. Ask Cryptography to locate it (it's
    # faster than looping in Python)
    try:
        ext = cert.extensions.get_extension_for_class(x509.SubjectAlternativeName).value
    except x509.ExtensionNotFound:
        # No such extension, return the empty list.
        return []
    except (
        x509.DuplicateExtension,
        UnsupportedExtension,
        x509.UnsupportedGeneralNameType,
        UnicodeError,
    ) as e:
        # A problem has been found with the quality of the certificate. Assume
        # no SAN field is present.
        log.warning(
            "A problem was encountered with the certificate that prevented "
            "urllib3 from finding the SubjectAlternativeName field. This can "
            "affect certificate validation. The error was %s",
            e,
        )
        return []

    # We want to return dNSName and iPAddress fields. We need to cast the IPs
    # back to strings because the match_hostname function wants them as
    # strings.
    # Sadly the DNS names need to be idna encoded and then, on Python 3, UTF-8
    # decoded. This is pretty frustrating, but that's what the standard library
    # does with certificates, and so we need to attempt to do the same.
    # We also want to skip over names which cannot be idna encoded.
    names = [
        ("DNS", name)
        for name in map(_dnsname_to_stdlib, ext.get_values_for_type(x509.DNSName))
        if name is not None
    ]
    names.extend(
        ("IP Address", str(name)) for name in ext.get_values_for_type(x509.IPAddress)
    )

    return names


class WrappedSocket(object):
    """API-compatibility wrapper for Python OpenSSL's Connection-class.

    Note: _makefile_refs, _drop() and _reuse() are needed for the garbage
    collector of pypy.
    """

    def __init__(self, connection, socket, suppress_ragged_eofs=True):
        self.connection = connection
        self.socket = socket
        self.suppress_ragged_eofs = suppress_ragged_eofs
        self._makefile_refs = 0
        self._closed = False

    def fileno(self):
        return self.socket.fileno()

    # Copy-pasted from Python 3.5 source code
    def _decref_socketios(self):
        if self._makefile_refs > 0:
            self._makefile_refs -= 1
        if self._closed:
            self.close()

    def recv(self, *args, **kwargs):
        try:
            data = self.connection.recv(*args, **kwargs)
        except OpenSSL.SSL.SysCallError as e:
            if self.suppress_ragged_eofs and e.args == (-1, "Unexpected EOF"):
                return b""
            else:
                raise SocketError(str(e))
        except OpenSSL.SSL.ZeroReturnError:
            if self.connection.get_shutdown() == OpenSSL.SSL.RECEIVED_SHUTDOWN:
                return b""
            else:
                raise
        except OpenSSL.SSL.WantReadError:
            if not util.wait_for_read(self.socket, self.socket.gettimeout()):
                raise timeout("The read operation timed out")
            else:
                return self.recv(*args, **kwargs)

        # TLS 1.3 post-handshake authentication
        except OpenSSL.SSL.Error as e:
            raise ssl.SSLError("read error: %r" % e)
        else:
            return data

    def recv_into(self, *args, **kwargs):
        try:
            return self.connection.recv_into(*args, **kwargs)
        except OpenSSL.SSL.SysCallError as e:
            if self.suppress_ragged_eofs and e.args == (-1, "Unexpected EOF"):
                return 0
            else:
                raise SocketError(str(e))
        except OpenSSL.SSL.ZeroReturnError:
            if self.connection.get_shutdown() == OpenSSL.SSL.RECEIVED_SHUTDOWN:
                return 0
            else:
                raise
        except OpenSSL.SSL.WantReadError:
            if not util.wait_for_read(self.socket, self.socket.gettimeout()):
                raise timeout("The read operation timed out")
            else:
                return self.recv_into(*args, **kwargs)

        # TLS 1.3 post-handshake authentication
        except OpenSSL.SSL.Error as e:
            raise ssl.SSLError("read error: %r" % e)

    def settimeout(self, timeout):
        return self.socket.settimeout(timeout)

    def _send_until_done(self, data):
        while True:
            try:
                return self.connection.send(data)
            except OpenSSL.SSL.WantWriteError:
                if not util.wait_for_write(self.socket, self.socket.gettimeout()):
                    raise timeout()
                continue
            except OpenSSL.SSL.SysCallError as e:
                raise SocketError(str(e))

    def sendall(self, data):
        total_sent = 0
        while total_sent < len(data):
            sent = self._send_until_done(
                data[total_sent : total_sent + SSL_WRITE_BLOCKSIZE]
            )
            total_sent += sent

    def shutdown(self):
        # FIXME rethrow compatible exceptions should we ever use this
        self.connection.shutdown()

    def close(self):
        if self._makefile_refs < 1:
            try:
                self._closed = True
                return self.connection.close()
            except OpenSSL.SSL.Error:
                return
        else:
            self._makefile_refs -= 1

    def getpeercert(self, binary_form=False):
        x509 = self.connection.get_peer_certificate()

        if not x509:
            return x509

        if binary_form:
            return OpenSSL.crypto.dump_certificate(OpenSSL.crypto.FILETYPE_ASN1, x509)

        return {
            "subject": ((("commonName", x509.get_subject().CN),),),
            "subjectAltName": get_subj_alt_name(x509),
        }

    def version(self):
        return self.connection.get_protocol_version_name()

    def _reuse(self):
        self._makefile_refs += 1

    def _drop(self):
        if self._makefile_refs < 1:
            self.close()
        else:
            self._makefile_refs -= 1


if _fileobject:  # Platform-specific: Python 2

    def makefile(self, mode, bufsize=-1):
        self._makefile_refs += 1
        return _fileobject(self, mode, bufsize, close=True)

else:  # Platform-specific: Python 3
    makefile = backport_makefile

WrappedSocket.makefile = makefile


class PyOpenSSLContext(object):
    """
    I am a wrapper class for the PyOpenSSL ``Context`` object. I am responsible
    for translating the interface of the standard library ``SSLContext`` object
    to calls into PyOpenSSL.
    """

    def __init__(self, protocol):
        self.protocol = _openssl_versions[protocol]
        self._ctx = OpenSSL.SSL.Context(self.protocol)
        self._options = 0
        self.check_hostname = False

    @property
    def options(self):
        return self._options

    @options.setter
    def options(self, value):
        self._options = value
        self._ctx.set_options(value)

    @property
    def verify_mode(self):
        return _openssl_to_stdlib_verify[self._ctx.get_verify_mode()]

    @verify_mode.setter
    def verify_mode(self, value):
        self._ctx.set_verify(_stdlib_to_openssl_verify[value], _verify_callback)

    def set_default_verify_paths(self):
        self._ctx.set_default_verify_paths()

    def set_ciphers(self, ciphers):
        if isinstance(ciphers, six.text_type):
            ciphers = ciphers.encode("utf-8")
        self._ctx.set_cipher_list(ciphers)

    def load_verify_locations(self, cafile=None, capath=None, cadata=None):
        if cafile is not None:
            cafile = cafile.encode("utf-8")
        if capath is not None:
            capath = capath.encode("utf-8")
        try:
            self._ctx.load_verify_locations(cafile, capath)
            if cadata is not None:
                self._ctx.load_verify_locations(BytesIO(cadata))
        except OpenSSL.SSL.Error as e:
            raise ssl.SSLError("unable to load trusted certificates: %r" % e)

    def load_cert_chain(self, certfile, keyfile=None, password=None):
        self._ctx.use_certificate_chain_file(certfile)
        if password is not None:
            if not isinstance(password, six.binary_type):
                password = password.encode("utf-8")
            self._ctx.set_passwd_cb(lambda *_: password)
        self._ctx.use_privatekey_file(keyfile or certfile)

    def set_alpn_protocols(self, protocols):
        protocols = [six.ensure_binary(p) for p in protocols]
        return self._ctx.set_alpn_protos(protocols)

    def wrap_socket(
        self,
        sock,
        server_side=False,
        do_handshake_on_connect=True,
        suppress_ragged_eofs=True,
        server_hostname=None,
    ):
        cnx = OpenSSL.SSL.Connection(self._ctx, sock)

        if isinstance(server_hostname, six.text_type):  # Platform-specific: Python 3
            server_hostname = server_hostname.encode("utf-8")

        if server_hostname is not None:
            cnx.set_tlsext_host_name(server_hostname)

        cnx.set_connect_state()

        while True:
            try:
                cnx.do_handshake()
            except OpenSSL.SSL.WantReadError:
                if not util.wait_for_read(sock, sock.gettimeout()):
                    raise timeout("select timed out")
                continue
            except OpenSSL.SSL.Error as e:
                raise ssl.SSLError("bad handshake: %r" % e)
            break

        return WrappedSocket(cnx, sock)


def _verify_callback(cnx, x509, err_no, err_depth, return_code):
    return err_no == 0

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\testing\assertsql.py ===
# testing/assertsql.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors


from __future__ import annotations

import collections
import contextlib
import itertools
import re

from .. import event
from ..engine import url
from ..engine.default import DefaultDialect
from ..schema import BaseDDLElement


class AssertRule:
    is_consumed = False
    errormessage = None
    consume_statement = True

    def process_statement(self, execute_observed):
        pass

    def no_more_statements(self):
        assert False, (
            "All statements are complete, but pending "
            "assertion rules remain"
        )


class SQLMatchRule(AssertRule):
    pass


class CursorSQL(SQLMatchRule):
    def __init__(self, statement, params=None, consume_statement=True):
        self.statement = statement
        self.params = params
        self.consume_statement = consume_statement

    def process_statement(self, execute_observed):
        stmt = execute_observed.statements[0]
        if self.statement != stmt.statement or (
            self.params is not None and self.params != stmt.parameters
        ):
            self.consume_statement = True
            self.errormessage = (
                "Testing for exact SQL %s parameters %s received %s %s"
                % (
                    self.statement,
                    self.params,
                    stmt.statement,
                    stmt.parameters,
                )
            )
        else:
            execute_observed.statements.pop(0)
            self.is_consumed = True
            if not execute_observed.statements:
                self.consume_statement = True


class CompiledSQL(SQLMatchRule):
    def __init__(
        self, statement, params=None, dialect="default", enable_returning=True
    ):
        self.statement = statement
        self.params = params
        self.dialect = dialect
        self.enable_returning = enable_returning

    def _compare_sql(self, execute_observed, received_statement):
        stmt = re.sub(r"[\n\t]", "", self.statement)
        return received_statement == stmt

    def _compile_dialect(self, execute_observed):
        if self.dialect == "default":
            dialect = DefaultDialect()
            # this is currently what tests are expecting
            # dialect.supports_default_values = True
            dialect.supports_default_metavalue = True

            if self.enable_returning:
                dialect.insert_returning = dialect.update_returning = (
                    dialect.delete_returning
                ) = True
                dialect.use_insertmanyvalues = True
                dialect.supports_multivalues_insert = True
                dialect.update_returning_multifrom = True
                dialect.delete_returning_multifrom = True
                # dialect.favor_returning_over_lastrowid = True
                # dialect.insert_null_pk_still_autoincrements = True

                # this is calculated but we need it to be True for this
                # to look like all the current RETURNING dialects
                assert dialect.insert_executemany_returning

            return dialect
        else:
            return url.URL.create(self.dialect).get_dialect()()

    def _received_statement(self, execute_observed):
        """reconstruct the statement and params in terms
        of a target dialect, which for CompiledSQL is just DefaultDialect."""

        context = execute_observed.context
        compare_dialect = self._compile_dialect(execute_observed)

        # received_statement runs a full compile().  we should not need to
        # consider extracted_parameters; if we do this indicates some state
        # is being sent from a previous cached query, which some misbehaviors
        # in the ORM can cause, see #6881
        cache_key = None  # execute_observed.context.compiled.cache_key
        extracted_parameters = (
            None  # execute_observed.context.extracted_parameters
        )

        if "schema_translate_map" in context.execution_options:
            map_ = context.execution_options["schema_translate_map"]
        else:
            map_ = None

        if isinstance(execute_observed.clauseelement, BaseDDLElement):
            compiled = execute_observed.clauseelement.compile(
                dialect=compare_dialect,
                schema_translate_map=map_,
            )
        else:
            compiled = execute_observed.clauseelement.compile(
                cache_key=cache_key,
                dialect=compare_dialect,
                column_keys=context.compiled.column_keys,
                for_executemany=context.compiled.for_executemany,
                schema_translate_map=map_,
            )
        _received_statement = re.sub(r"[\n\t]", "", str(compiled))
        parameters = execute_observed.parameters

        if not parameters:
            _received_parameters = [
                compiled.construct_params(
                    extracted_parameters=extracted_parameters
                )
            ]
        else:
            _received_parameters = [
                compiled.construct_params(
                    m, extracted_parameters=extracted_parameters
                )
                for m in parameters
            ]

        return _received_statement, _received_parameters

    def process_statement(self, execute_observed):
        context = execute_observed.context

        _received_statement, _received_parameters = self._received_statement(
            execute_observed
        )
        params = self._all_params(context)

        equivalent = self._compare_sql(execute_observed, _received_statement)

        if equivalent:
            if params is not None:
                all_params = list(params)
                all_received = list(_received_parameters)
                while all_params and all_received:
                    param = dict(all_params.pop(0))

                    for idx, received in enumerate(list(all_received)):
                        # do a positive compare only
                        for param_key in param:
                            # a key in param did not match current
                            # 'received'
                            if (
                                param_key not in received
                                or received[param_key] != param[param_key]
                            ):
                                break
                        else:
                            # all keys in param matched 'received';
                            # onto next param
                            del all_received[idx]
                            break
                    else:
                        # param did not match any entry
                        # in all_received
                        equivalent = False
                        break
                if all_params or all_received:
                    equivalent = False

        if equivalent:
            self.is_consumed = True
            self.errormessage = None
        else:
            self.errormessage = self._failure_message(
                execute_observed, params
            ) % {
                "received_statement": _received_statement,
                "received_parameters": _received_parameters,
            }

    def _all_params(self, context):
        if self.params:
            if callable(self.params):
                params = self.params(context)
            else:
                params = self.params
            if not isinstance(params, list):
                params = [params]
            return params
        else:
            return None

    def _failure_message(self, execute_observed, expected_params):
        return (
            "Testing for compiled statement\n%r partial params %s, "
            "received\n%%(received_statement)r with params "
            "%%(received_parameters)r"
            % (
                self.statement.replace("%", "%%"),
                repr(expected_params).replace("%", "%%"),
            )
        )


class RegexSQL(CompiledSQL):
    def __init__(
        self, regex, params=None, dialect="default", enable_returning=False
    ):
        SQLMatchRule.__init__(self)
        self.regex = re.compile(regex)
        self.orig_regex = regex
        self.params = params
        self.dialect = dialect
        self.enable_returning = enable_returning

    def _failure_message(self, execute_observed, expected_params):
        return (
            "Testing for compiled statement ~%r partial params %s, "
            "received %%(received_statement)r with params "
            "%%(received_parameters)r"
            % (
                self.orig_regex.replace("%", "%%"),
                repr(expected_params).replace("%", "%%"),
            )
        )

    def _compare_sql(self, execute_observed, received_statement):
        return bool(self.regex.match(received_statement))


class DialectSQL(CompiledSQL):
    def _compile_dialect(self, execute_observed):
        return execute_observed.context.dialect

    def _compare_no_space(self, real_stmt, received_stmt):
        stmt = re.sub(r"[\n\t]", "", real_stmt)
        return received_stmt == stmt

    def _received_statement(self, execute_observed):
        received_stmt, received_params = super()._received_statement(
            execute_observed
        )

        # TODO: why do we need this part?
        for real_stmt in execute_observed.statements:
            if self._compare_no_space(real_stmt.statement, received_stmt):
                break
        else:
            raise AssertionError(
                "Can't locate compiled statement %r in list of "
                "statements actually invoked" % received_stmt
            )

        return received_stmt, execute_observed.context.compiled_parameters

    def _dialect_adjusted_statement(self, dialect):
        paramstyle = dialect.paramstyle
        stmt = re.sub(r"[\n\t]", "", self.statement)

        # temporarily escape out PG double colons
        stmt = stmt.replace("::", "!!")

        if paramstyle == "pyformat":
            stmt = re.sub(r":([\w_]+)", r"%(\1)s", stmt)
        else:
            # positional params
            repl = None
            if paramstyle == "qmark":
                repl = "?"
            elif paramstyle == "format":
                repl = r"%s"
            elif paramstyle.startswith("numeric"):
                counter = itertools.count(1)

                num_identifier = "$" if paramstyle == "numeric_dollar" else ":"

                def repl(m):
                    return f"{num_identifier}{next(counter)}"

            stmt = re.sub(r":([\w_]+)", repl, stmt)

        # put them back
        stmt = stmt.replace("!!", "::")

        return stmt

    def _compare_sql(self, execute_observed, received_statement):
        stmt = self._dialect_adjusted_statement(
            execute_observed.context.dialect
        )
        return received_statement == stmt

    def _failure_message(self, execute_observed, expected_params):
        return (
            "Testing for compiled statement\n%r partial params %s, "
            "received\n%%(received_statement)r with params "
            "%%(received_parameters)r"
            % (
                self._dialect_adjusted_statement(
                    execute_observed.context.dialect
                ).replace("%", "%%"),
                repr(expected_params).replace("%", "%%"),
            )
        )


class CountStatements(AssertRule):
    def __init__(self, count):
        self.count = count
        self._statement_count = 0

    def process_statement(self, execute_observed):
        self._statement_count += 1

    def no_more_statements(self):
        if self.count != self._statement_count:
            assert False, "desired statement count %d does not match %d" % (
                self.count,
                self._statement_count,
            )


class AllOf(AssertRule):
    def __init__(self, *rules):
        self.rules = set(rules)

    def process_statement(self, execute_observed):
        for rule in list(self.rules):
            rule.errormessage = None
            rule.process_statement(execute_observed)
            if rule.is_consumed:
                self.rules.discard(rule)
                if not self.rules:
                    self.is_consumed = True
                break
            elif not rule.errormessage:
                # rule is not done yet
                self.errormessage = None
                break
        else:
            self.errormessage = list(self.rules)[0].errormessage


class EachOf(AssertRule):
    def __init__(self, *rules):
        self.rules = list(rules)

    def process_statement(self, execute_observed):
        if not self.rules:
            self.is_consumed = True
            self.consume_statement = False

        while self.rules:
            rule = self.rules[0]
            rule.process_statement(execute_observed)
            if rule.is_consumed:
                self.rules.pop(0)
            elif rule.errormessage:
                self.errormessage = rule.errormessage
            if rule.consume_statement:
                break

        if not self.rules:
            self.is_consumed = True

    def no_more_statements(self):
        if self.rules and not self.rules[0].is_consumed:
            self.rules[0].no_more_statements()
        elif self.rules:
            super().no_more_statements()


class Conditional(EachOf):
    def __init__(self, condition, rules, else_rules):
        if condition:
            super().__init__(*rules)
        else:
            super().__init__(*else_rules)


class Or(AllOf):
    def process_statement(self, execute_observed):
        for rule in self.rules:
            rule.process_statement(execute_observed)
            if rule.is_consumed:
                self.is_consumed = True
                break
        else:
            self.errormessage = list(self.rules)[0].errormessage


class SQLExecuteObserved:
    def __init__(self, context, clauseelement, multiparams, params):
        self.context = context
        self.clauseelement = clauseelement

        if multiparams:
            self.parameters = multiparams
        elif params:
            self.parameters = [params]
        else:
            self.parameters = []
        self.statements = []

    def __repr__(self):
        return str(self.statements)


class SQLCursorExecuteObserved(
    collections.namedtuple(
        "SQLCursorExecuteObserved",
        ["statement", "parameters", "context", "executemany"],
    )
):
    pass


class SQLAsserter:
    def __init__(self):
        self.accumulated = []

    def _close(self):
        self._final = self.accumulated
        del self.accumulated

    def assert_(self, *rules):
        rule = EachOf(*rules)

        observed = list(self._final)
        while observed:
            statement = observed.pop(0)
            rule.process_statement(statement)
            if rule.is_consumed:
                break
            elif rule.errormessage:
                assert False, rule.errormessage
        if observed:
            assert False, "Additional SQL statements remain:\n%s" % observed
        elif not rule.is_consumed:
            rule.no_more_statements()


@contextlib.contextmanager
def assert_engine(engine):
    asserter = SQLAsserter()

    orig = []

    @event.listens_for(engine, "before_execute")
    def connection_execute(
        conn, clauseelement, multiparams, params, execution_options
    ):
        # grab the original statement + params before any cursor
        # execution
        orig[:] = clauseelement, multiparams, params

    @event.listens_for(engine, "after_cursor_execute")
    def cursor_execute(
        conn, cursor, statement, parameters, context, executemany
    ):
        if not context:
            return
        # then grab real cursor statements and associate them all
        # around a single context
        if (
            asserter.accumulated
            and asserter.accumulated[-1].context is context
        ):
            obs = asserter.accumulated[-1]
        else:
            obs = SQLExecuteObserved(context, orig[0], orig[1], orig[2])
            asserter.accumulated.append(obs)
        obs.statements.append(
            SQLCursorExecuteObserved(
                statement, parameters, context, executemany
            )
        )

    try:
        yield asserter
    finally:
        event.remove(engine, "after_cursor_execute", cursor_execute)
        event.remove(engine, "before_execute", connection_execute)
        asserter._close()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\numberfields\subfield.py ===
r"""
Functions in ``polys.numberfields.subfield`` solve the "Subfield Problem" and
allied problems, for algebraic number fields.

Following Cohen (see [Cohen93]_ Section 4.5), we can define the main problem as
follows:

* **Subfield Problem:**

  Given two number fields $\mathbb{Q}(\alpha)$, $\mathbb{Q}(\beta)$
  via the minimal polynomials for their generators $\alpha$ and $\beta$, decide
  whether one field is isomorphic to a subfield of the other.

From a solution to this problem flow solutions to the following problems as
well:

* **Primitive Element Problem:**

  Given several algebraic numbers
  $\alpha_1, \ldots, \alpha_m$, compute a single algebraic number $\theta$
  such that $\mathbb{Q}(\alpha_1, \ldots, \alpha_m) = \mathbb{Q}(\theta)$.

* **Field Isomorphism Problem:**

  Decide whether two number fields
  $\mathbb{Q}(\alpha)$, $\mathbb{Q}(\beta)$ are isomorphic.

* **Field Membership Problem:**

  Given two algebraic numbers $\alpha$,
  $\beta$, decide whether $\alpha \in \mathbb{Q}(\beta)$, and if so write
  $\alpha = f(\beta)$ for some $f(x) \in \mathbb{Q}[x]$.
"""

from sympy.core.add import Add
from sympy.core.numbers import AlgebraicNumber
from sympy.core.singleton import S
from sympy.core.symbol import Dummy
from sympy.core.sympify import sympify, _sympify
from sympy.ntheory import sieve
from sympy.polys.densetools import dup_eval
from sympy.polys.domains import QQ
from sympy.polys.numberfields.minpoly import _choose_factor, minimal_polynomial
from sympy.polys.polyerrors import IsomorphismFailed
from sympy.polys.polytools import Poly, PurePoly, factor_list
from sympy.utilities import public

from mpmath import MPContext


def is_isomorphism_possible(a, b):
    """Necessary but not sufficient test for isomorphism. """
    n = a.minpoly.degree()
    m = b.minpoly.degree()

    if m % n != 0:
        return False

    if n == m:
        return True

    da = a.minpoly.discriminant()
    db = b.minpoly.discriminant()

    i, k, half = 1, m//n, db//2

    while True:
        p = sieve[i]
        P = p**k

        if P > half:
            break

        if ((da % p) % 2) and not (db % P):
            return False

        i += 1

    return True


def field_isomorphism_pslq(a, b):
    """Construct field isomorphism using PSLQ algorithm. """
    if not a.root.is_real or not b.root.is_real:
        raise NotImplementedError("PSLQ doesn't support complex coefficients")

    f = a.minpoly
    g = b.minpoly.replace(f.gen)

    n, m, prev = 100, b.minpoly.degree(), None
    ctx = MPContext()

    for i in range(1, 5):
        A = a.root.evalf(n)
        B = b.root.evalf(n)

        basis = [1, B] + [ B**i for i in range(2, m) ] + [-A]

        ctx.dps = n
        coeffs = ctx.pslq(basis, maxcoeff=10**10, maxsteps=1000)

        if coeffs is None:
            # PSLQ can't find an integer linear combination. Give up.
            break

        if coeffs != prev:
            prev = coeffs
        else:
            # Increasing precision didn't produce anything new. Give up.
            break

        # We have
        #   c0 + c1*B + c2*B^2 + ... + cm-1*B^(m-1) - cm*A ~ 0.
        # So bring cm*A to the other side, and divide through by cm,
        # for an approximate representation of A as a polynomial in B.
        # (We know cm != 0 since `b.minpoly` is irreducible.)
        coeffs = [S(c)/coeffs[-1] for c in coeffs[:-1]]

        # Throw away leading zeros.
        while not coeffs[-1]:
            coeffs.pop()

        coeffs = list(reversed(coeffs))
        h = Poly(coeffs, f.gen, domain='QQ')

        # We only have A ~ h(B). We must check whether the relation is exact.
        if f.compose(h).rem(g).is_zero:
            # Now we know that h(b) is in fact equal to _some conjugate of_ a.
            # But from the very precise approximation A ~ h(B) we can assume
            # the conjugate is a itself.
            return coeffs
        else:
            n *= 2

    return None


def field_isomorphism_factor(a, b):
    """Construct field isomorphism via factorization. """
    _, factors = factor_list(a.minpoly, extension=b)
    for f, _ in factors:
        if f.degree() == 1:
            # Any linear factor f(x) represents some conjugate of a in QQ(b).
            # We want to know whether this linear factor represents a itself.
            # Let f = x - c
            c = -f.rep.TC()
            # Write c as polynomial in b
            coeffs = c.to_sympy_list()
            d, terms = len(coeffs) - 1, []
            for i, coeff in enumerate(coeffs):
                terms.append(coeff*b.root**(d - i))
            r = Add(*terms)
            # Check whether we got the number a
            if a.minpoly.same_root(r, a):
                return coeffs

    # If none of the linear factors represented a in QQ(b), then in fact a is
    # not an element of QQ(b).
    return None


@public
def field_isomorphism(a, b, *, fast=True):
    r"""
    Find an embedding of one number field into another.

    Explanation
    ===========

    This function looks for an isomorphism from $\mathbb{Q}(a)$ onto some
    subfield of $\mathbb{Q}(b)$. Thus, it solves the Subfield Problem.

    Examples
    ========

    >>> from sympy import sqrt, field_isomorphism, I
    >>> print(field_isomorphism(3, sqrt(2)))  # doctest: +SKIP
    [3]
    >>> print(field_isomorphism( I*sqrt(3), I*sqrt(3)/2))  # doctest: +SKIP
    [2, 0]

    Parameters
    ==========

    a : :py:class:`~.Expr`
        Any expression representing an algebraic number.
    b : :py:class:`~.Expr`
        Any expression representing an algebraic number.
    fast : boolean, optional (default=True)
        If ``True``, we first attempt a potentially faster way of computing the
        isomorphism, falling back on a slower method if this fails. If
        ``False``, we go directly to the slower method, which is guaranteed to
        return a result.

    Returns
    =======

    List of rational numbers, or None
        If $\mathbb{Q}(a)$ is not isomorphic to some subfield of
        $\mathbb{Q}(b)$, then return ``None``. Otherwise, return a list of
        rational numbers representing an element of $\mathbb{Q}(b)$ to which
        $a$ may be mapped, in order to define a monomorphism, i.e. an
        isomorphism from $\mathbb{Q}(a)$ to some subfield of $\mathbb{Q}(b)$.
        The elements of the list are the coefficients of falling powers of $b$.

    """
    a, b = sympify(a), sympify(b)

    if not a.is_AlgebraicNumber:
        a = AlgebraicNumber(a)

    if not b.is_AlgebraicNumber:
        b = AlgebraicNumber(b)

    a = a.to_primitive_element()
    b = b.to_primitive_element()

    if a == b:
        return a.coeffs()

    n = a.minpoly.degree()
    m = b.minpoly.degree()

    if n == 1:
        return [a.root]

    if m % n != 0:
        return None

    if fast:
        try:
            result = field_isomorphism_pslq(a, b)

            if result is not None:
                return result
        except NotImplementedError:
            pass

    return field_isomorphism_factor(a, b)


def _switch_domain(g, K):
    # An algebraic relation f(a, b) = 0 over Q can also be written
    # g(b) = 0 where g is in Q(a)[x] and h(a) = 0 where h is in Q(b)[x].
    # This function transforms g into h where Q(b) = K.
    frep = g.rep.inject()
    hrep = frep.eject(K, front=True)

    return g.new(hrep, g.gens[0])


def _linsolve(p):
    # Compute root of linear polynomial.
    c, d = p.rep.to_list()
    return -d/c


@public
def primitive_element(extension, x=None, *, ex=False, polys=False):
    r"""
    Find a single generator for a number field given by several generators.

    Explanation
    ===========

    The basic problem is this: Given several algebraic numbers
    $\alpha_1, \alpha_2, \ldots, \alpha_n$, find a single algebraic number
    $\theta$ such that
    $\mathbb{Q}(\alpha_1, \alpha_2, \ldots, \alpha_n) = \mathbb{Q}(\theta)$.

    This function actually guarantees that $\theta$ will be a linear
    combination of the $\alpha_i$, with non-negative integer coefficients.

    Furthermore, if desired, this function will tell you how to express each
    $\alpha_i$ as a $\mathbb{Q}$-linear combination of the powers of $\theta$.

    Examples
    ========

    >>> from sympy import primitive_element, sqrt, S, minpoly, simplify
    >>> from sympy.abc import x
    >>> f, lincomb, reps = primitive_element([sqrt(2), sqrt(3)], x, ex=True)

    Then ``lincomb`` tells us the primitive element as a linear combination of
    the given generators ``sqrt(2)`` and ``sqrt(3)``.

    >>> print(lincomb)
    [1, 1]

    This means the primtiive element is $\sqrt{2} + \sqrt{3}$.
    Meanwhile ``f`` is the minimal polynomial for this primitive element.

    >>> print(f)
    x**4 - 10*x**2 + 1
    >>> print(minpoly(sqrt(2) + sqrt(3), x))
    x**4 - 10*x**2 + 1

    Finally, ``reps`` (which was returned only because we set keyword arg
    ``ex=True``) tells us how to recover each of the generators $\sqrt{2}$ and
    $\sqrt{3}$ as $\mathbb{Q}$-linear combinations of the powers of the
    primitive element $\sqrt{2} + \sqrt{3}$.

    >>> print([S(r) for r in reps[0]])
    [1/2, 0, -9/2, 0]
    >>> theta = sqrt(2) + sqrt(3)
    >>> print(simplify(theta**3/2 - 9*theta/2))
    sqrt(2)
    >>> print([S(r) for r in reps[1]])
    [-1/2, 0, 11/2, 0]
    >>> print(simplify(-theta**3/2 + 11*theta/2))
    sqrt(3)

    Parameters
    ==========

    extension : list of :py:class:`~.Expr`
        Each expression must represent an algebraic number $\alpha_i$.
    x : :py:class:`~.Symbol`, optional (default=None)
        The desired symbol to appear in the computed minimal polynomial for the
        primitive element $\theta$. If ``None``, we use a dummy symbol.
    ex : boolean, optional (default=False)
        If and only if ``True``, compute the representation of each $\alpha_i$
        as a $\mathbb{Q}$-linear combination over the powers of $\theta$.
    polys : boolean, optional (default=False)
        If ``True``, return the minimal polynomial as a :py:class:`~.Poly`.
        Otherwise return it as an :py:class:`~.Expr`.

    Returns
    =======

    Pair (f, coeffs) or triple (f, coeffs, reps), where:
        ``f`` is the minimal polynomial for the primitive element.
        ``coeffs`` gives the primitive element as a linear combination of the
        given generators.
        ``reps`` is present if and only if argument ``ex=True`` was passed,
        and is a list of lists of rational numbers. Each list gives the
        coefficients of falling powers of the primitive element, to recover
        one of the original, given generators.

    """
    if not extension:
        raise ValueError("Cannot compute primitive element for empty extension")
    extension = [_sympify(ext) for ext in extension]

    if x is not None:
        x, cls = sympify(x), Poly
    else:
        x, cls = Dummy('x'), PurePoly

    def _canonicalize(f):
        _, f = f.primitive()
        if f.LC() < 0:
            f = -f
        return f

    if not ex:
        gen, coeffs = extension[0], [1]
        g = minimal_polynomial(gen, x, polys=True)
        for ext in extension[1:]:
            if ext.is_Rational:
                coeffs.append(0)
                continue
            _, factors = factor_list(g, extension=ext)
            g = _choose_factor(factors, x, gen)
            [s], _, g = g.sqf_norm()
            gen += s*ext
            coeffs.append(s)

        g = _canonicalize(g)
        if not polys:
            return g.as_expr(), coeffs
        else:
            return cls(g), coeffs

    gen, coeffs = extension[0], [1]
    f = minimal_polynomial(gen, x, polys=True)
    K = QQ.algebraic_field((f, gen))  # incrementally constructed field
    reps = [K.unit]  # representations of extension elements in K
    for ext in extension[1:]:
        if ext.is_Rational:
            coeffs.append(0)    # rational ext is not included in the expression of a primitive element
            reps.append(K.convert(ext))    # but it is included in reps
            continue
        p = minimal_polynomial(ext, x, polys=True)
        L = QQ.algebraic_field((p, ext))
        _, factors = factor_list(f, domain=L)
        f = _choose_factor(factors, x, gen)
        [s], g, f = f.sqf_norm()
        gen += s*ext
        coeffs.append(s)
        K = QQ.algebraic_field((f, gen))
        h = _switch_domain(g, K)
        erep = _linsolve(h.gcd(p))  # ext as element of K
        ogen = K.unit - s*erep  # old gen as element of K
        reps = [dup_eval(_.to_list(), ogen, K) for _ in reps] + [erep]

    if K.ext.root.is_Rational:  # all extensions are rational
        H = [K.convert(_).rep for _ in extension]
        coeffs = [0]*len(extension)
        f = cls(x, domain=QQ)
    else:
        H = [_.to_list() for _ in reps]

    f = _canonicalize(f)
    if not polys:
        return f.as_expr(), coeffs, H
    else:
        return f, coeffs, H


@public
def to_number_field(extension, theta=None, *, gen=None, alias=None):
    r"""
    Express one algebraic number in the field generated by another.

    Explanation
    ===========

    Given two algebraic numbers $\eta, \theta$, this function either expresses
    $\eta$ as an element of $\mathbb{Q}(\theta)$, or else raises an exception
    if $\eta \not\in \mathbb{Q}(\theta)$.

    This function is essentially just a convenience, utilizing
    :py:func:`~.field_isomorphism` (our solution of the Subfield Problem) to
    solve this, the Field Membership Problem.

    As an additional convenience, this function allows you to pass a list of
    algebraic numbers $\alpha_1, \alpha_2, \ldots, \alpha_n$ instead of $\eta$.
    It then computes $\eta$ for you, as a solution of the Primitive Element
    Problem, using :py:func:`~.primitive_element` on the list of $\alpha_i$.

    Examples
    ========

    >>> from sympy import sqrt, to_number_field
    >>> eta = sqrt(2)
    >>> theta = sqrt(2) + sqrt(3)
    >>> a = to_number_field(eta, theta)
    >>> print(type(a))
    <class 'sympy.core.numbers.AlgebraicNumber'>
    >>> a.root
    sqrt(2) + sqrt(3)
    >>> print(a)
    sqrt(2)
    >>> a.coeffs()
    [1/2, 0, -9/2, 0]

    We get an :py:class:`~.AlgebraicNumber`, whose ``.root`` is $\theta$, whose
    value is $\eta$, and whose ``.coeffs()`` show how to write $\eta$ as a
    $\mathbb{Q}$-linear combination in falling powers of $\theta$.

    Parameters
    ==========

    extension : :py:class:`~.Expr` or list of :py:class:`~.Expr`
        Either the algebraic number that is to be expressed in the other field,
        or else a list of algebraic numbers, a primitive element for which is
        to be expressed in the other field.
    theta : :py:class:`~.Expr`, None, optional (default=None)
        If an :py:class:`~.Expr` representing an algebraic number, behavior is
        as described under **Explanation**. If ``None``, then this function
        reduces to a shorthand for calling :py:func:`~.primitive_element` on
        ``extension`` and turning the computed primitive element into an
        :py:class:`~.AlgebraicNumber`.
    gen : :py:class:`~.Symbol`, None, optional (default=None)
        If provided, this will be used as the generator symbol for the minimal
        polynomial in the returned :py:class:`~.AlgebraicNumber`.
    alias : str, :py:class:`~.Symbol`, None, optional (default=None)
        If provided, this will be used as the alias symbol for the returned
        :py:class:`~.AlgebraicNumber`.

    Returns
    =======

    AlgebraicNumber
        Belonging to $\mathbb{Q}(\theta)$ and equaling $\eta$.

    Raises
    ======

    IsomorphismFailed
        If $\eta \not\in \mathbb{Q}(\theta)$.

    See Also
    ========

    field_isomorphism
    primitive_element

    """
    if hasattr(extension, '__iter__'):
        extension = list(extension)
    else:
        extension = [extension]

    if len(extension) == 1 and isinstance(extension[0], tuple):
        return AlgebraicNumber(extension[0], alias=alias)

    minpoly, coeffs = primitive_element(extension, gen, polys=True)
    root = sum(coeff*ext for coeff, ext in zip(coeffs, extension))

    if theta is None:
        return AlgebraicNumber((minpoly, root), alias=alias)
    else:
        theta = sympify(theta)

        if not theta.is_AlgebraicNumber:
            theta = AlgebraicNumber(theta, gen=gen, alias=alias)

        coeffs = field_isomorphism(root, theta)

        if coeffs is not None:
            return AlgebraicNumber(theta, coeffs, alias=alias)
        else:
            raise IsomorphismFailed(
                "%s is not in a subfield of %s" % (root, theta.root))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\ext\mypy\decl_class.py ===
# ext/mypy/decl_class.py
# Copyright (C) 2021-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

from __future__ import annotations

from typing import List
from typing import Optional
from typing import Union

from mypy.nodes import AssignmentStmt
from mypy.nodes import CallExpr
from mypy.nodes import ClassDef
from mypy.nodes import Decorator
from mypy.nodes import LambdaExpr
from mypy.nodes import ListExpr
from mypy.nodes import MemberExpr
from mypy.nodes import NameExpr
from mypy.nodes import PlaceholderNode
from mypy.nodes import RefExpr
from mypy.nodes import StrExpr
from mypy.nodes import SymbolNode
from mypy.nodes import SymbolTableNode
from mypy.nodes import TempNode
from mypy.nodes import TypeInfo
from mypy.nodes import Var
from mypy.plugin import SemanticAnalyzerPluginInterface
from mypy.types import AnyType
from mypy.types import CallableType
from mypy.types import get_proper_type
from mypy.types import Instance
from mypy.types import NoneType
from mypy.types import ProperType
from mypy.types import Type
from mypy.types import TypeOfAny
from mypy.types import UnboundType
from mypy.types import UnionType

from . import apply
from . import infer
from . import names
from . import util


def scan_declarative_assignments_and_apply_types(
    cls: ClassDef,
    api: SemanticAnalyzerPluginInterface,
    is_mixin_scan: bool = False,
) -> Optional[List[util.SQLAlchemyAttribute]]:
    info = util.info_for_cls(cls, api)

    if info is None:
        # this can occur during cached passes
        return None
    elif cls.fullname.startswith("builtins"):
        return None

    mapped_attributes: Optional[List[util.SQLAlchemyAttribute]] = (
        util.get_mapped_attributes(info, api)
    )

    # used by assign.add_additional_orm_attributes among others
    util.establish_as_sqlalchemy(info)

    if mapped_attributes is not None:
        # ensure that a class that's mapped is always picked up by
        # its mapped() decorator or declarative metaclass before
        # it would be detected as an unmapped mixin class

        if not is_mixin_scan:
            # mypy can call us more than once.  it then *may* have reset the
            # left hand side of everything, but not the right that we removed,
            # removing our ability to re-scan.   but we have the types
            # here, so lets re-apply them, or if we have an UnboundType,
            # we can re-scan

            apply.re_apply_declarative_assignments(cls, api, mapped_attributes)

        return mapped_attributes

    mapped_attributes = []

    if not cls.defs.body:
        # when we get a mixin class from another file, the body is
        # empty (!) but the names are in the symbol table.  so use that.

        for sym_name, sym in info.names.items():
            _scan_symbol_table_entry(
                cls, api, sym_name, sym, mapped_attributes
            )
    else:
        for stmt in util.flatten_typechecking(cls.defs.body):
            if isinstance(stmt, AssignmentStmt):
                _scan_declarative_assignment_stmt(
                    cls, api, stmt, mapped_attributes
                )
            elif isinstance(stmt, Decorator):
                _scan_declarative_decorator_stmt(
                    cls, api, stmt, mapped_attributes
                )
    _scan_for_mapped_bases(cls, api)

    if not is_mixin_scan:
        apply.add_additional_orm_attributes(cls, api, mapped_attributes)

    util.set_mapped_attributes(info, mapped_attributes)

    return mapped_attributes


def _scan_symbol_table_entry(
    cls: ClassDef,
    api: SemanticAnalyzerPluginInterface,
    name: str,
    value: SymbolTableNode,
    attributes: List[util.SQLAlchemyAttribute],
) -> None:
    """Extract mapping information from a SymbolTableNode that's in the
    type.names dictionary.

    """
    value_type = get_proper_type(value.type)
    if not isinstance(value_type, Instance):
        return

    left_hand_explicit_type = None
    type_id = names.type_id_for_named_node(value_type.type)
    # type_id = names._type_id_for_unbound_type(value.type.type, cls, api)

    err = False

    # TODO: this is nearly the same logic as that of
    # _scan_declarative_decorator_stmt, likely can be merged
    if type_id in {
        names.MAPPED,
        names.RELATIONSHIP,
        names.COMPOSITE_PROPERTY,
        names.MAPPER_PROPERTY,
        names.SYNONYM_PROPERTY,
        names.COLUMN_PROPERTY,
    }:
        if value_type.args:
            left_hand_explicit_type = get_proper_type(value_type.args[0])
        else:
            err = True
    elif type_id is names.COLUMN:
        if not value_type.args:
            err = True
        else:
            typeengine_arg: Union[ProperType, TypeInfo] = get_proper_type(
                value_type.args[0]
            )
            if isinstance(typeengine_arg, Instance):
                typeengine_arg = typeengine_arg.type

            if isinstance(typeengine_arg, (UnboundType, TypeInfo)):
                sym = api.lookup_qualified(typeengine_arg.name, typeengine_arg)
                if sym is not None and isinstance(sym.node, TypeInfo):
                    if names.has_base_type_id(sym.node, names.TYPEENGINE):
                        left_hand_explicit_type = UnionType(
                            [
                                infer.extract_python_type_from_typeengine(
                                    api, sym.node, []
                                ),
                                NoneType(),
                            ]
                        )
                    else:
                        util.fail(
                            api,
                            "Column type should be a TypeEngine "
                            "subclass not '{}'".format(sym.node.fullname),
                            value_type,
                        )

    if err:
        msg = (
            "Can't infer type from attribute {} on class {}. "
            "please specify a return type from this function that is "
            "one of: Mapped[<python type>], relationship[<target class>], "
            "Column[<TypeEngine>], MapperProperty[<python type>]"
        )
        util.fail(api, msg.format(name, cls.name), cls)

        left_hand_explicit_type = AnyType(TypeOfAny.special_form)

    if left_hand_explicit_type is not None:
        assert value.node is not None
        attributes.append(
            util.SQLAlchemyAttribute(
                name=name,
                line=value.node.line,
                column=value.node.column,
                typ=left_hand_explicit_type,
                info=cls.info,
            )
        )


def _scan_declarative_decorator_stmt(
    cls: ClassDef,
    api: SemanticAnalyzerPluginInterface,
    stmt: Decorator,
    attributes: List[util.SQLAlchemyAttribute],
) -> None:
    """Extract mapping information from a @declared_attr in a declarative
    class.

    E.g.::

        @reg.mapped
        class MyClass:
            # ...

            @declared_attr
            def updated_at(cls) -> Column[DateTime]:
                return Column(DateTime)

    Will resolve in mypy as::

        @reg.mapped
        class MyClass:
            # ...

            updated_at: Mapped[Optional[datetime.datetime]]

    """
    for dec in stmt.decorators:
        if (
            isinstance(dec, (NameExpr, MemberExpr, SymbolNode))
            and names.type_id_for_named_node(dec) is names.DECLARED_ATTR
        ):
            break
    else:
        return

    dec_index = cls.defs.body.index(stmt)

    left_hand_explicit_type: Optional[ProperType] = None

    if util.name_is_dunder(stmt.name):
        # for dunder names like __table_args__, __tablename__,
        # __mapper_args__ etc., rewrite these as simple assignment
        # statements; otherwise mypy doesn't like if the decorated
        # function has an annotation like ``cls: Type[Foo]`` because
        # it isn't @classmethod
        any_ = AnyType(TypeOfAny.special_form)
        left_node = NameExpr(stmt.var.name)
        left_node.node = stmt.var
        new_stmt = AssignmentStmt([left_node], TempNode(any_))
        new_stmt.type = left_node.node.type
        cls.defs.body[dec_index] = new_stmt
        return
    elif isinstance(stmt.func.type, CallableType):
        func_type = stmt.func.type.ret_type
        if isinstance(func_type, UnboundType):
            type_id = names.type_id_for_unbound_type(func_type, cls, api)
        else:
            # this does not seem to occur unless the type argument is
            # incorrect
            return

        if (
            type_id
            in {
                names.MAPPED,
                names.RELATIONSHIP,
                names.COMPOSITE_PROPERTY,
                names.MAPPER_PROPERTY,
                names.SYNONYM_PROPERTY,
                names.COLUMN_PROPERTY,
            }
            and func_type.args
        ):
            left_hand_explicit_type = get_proper_type(func_type.args[0])
        elif type_id is names.COLUMN and func_type.args:
            typeengine_arg = func_type.args[0]
            if isinstance(typeengine_arg, UnboundType):
                sym = api.lookup_qualified(typeengine_arg.name, typeengine_arg)
                if sym is not None and isinstance(sym.node, TypeInfo):
                    if names.has_base_type_id(sym.node, names.TYPEENGINE):
                        left_hand_explicit_type = UnionType(
                            [
                                infer.extract_python_type_from_typeengine(
                                    api, sym.node, []
                                ),
                                NoneType(),
                            ]
                        )
                    else:
                        util.fail(
                            api,
                            "Column type should be a TypeEngine "
                            "subclass not '{}'".format(sym.node.fullname),
                            func_type,
                        )

    if left_hand_explicit_type is None:
        # no type on the decorated function.  our option here is to
        # dig into the function body and get the return type, but they
        # should just have an annotation.
        msg = (
            "Can't infer type from @declared_attr on function '{}';  "
            "please specify a return type from this function that is "
            "one of: Mapped[<python type>], relationship[<target class>], "
            "Column[<TypeEngine>], MapperProperty[<python type>]"
        )
        util.fail(api, msg.format(stmt.var.name), stmt)

        left_hand_explicit_type = AnyType(TypeOfAny.special_form)

    left_node = NameExpr(stmt.var.name)
    left_node.node = stmt.var

    # totally feeling around in the dark here as I don't totally understand
    # the significance of UnboundType.  It seems to be something that is
    # not going to do what's expected when it is applied as the type of
    # an AssignmentStatement.  So do a feeling-around-in-the-dark version
    # of converting it to the regular Instance/TypeInfo/UnionType structures
    # we see everywhere else.
    if isinstance(left_hand_explicit_type, UnboundType):
        left_hand_explicit_type = get_proper_type(
            util.unbound_to_instance(api, left_hand_explicit_type)
        )

    left_node.node.type = api.named_type(
        names.NAMED_TYPE_SQLA_MAPPED, [left_hand_explicit_type]
    )

    # this will ignore the rvalue entirely
    # rvalue = TempNode(AnyType(TypeOfAny.special_form))

    # rewrite the node as:
    # <attr> : Mapped[<typ>] =
    # _sa_Mapped._empty_constructor(lambda: <function body>)
    # the function body is maintained so it gets type checked internally
    rvalue = names.expr_to_mapped_constructor(
        LambdaExpr(stmt.func.arguments, stmt.func.body)
    )

    new_stmt = AssignmentStmt([left_node], rvalue)
    new_stmt.type = left_node.node.type

    attributes.append(
        util.SQLAlchemyAttribute(
            name=left_node.name,
            line=stmt.line,
            column=stmt.column,
            typ=left_hand_explicit_type,
            info=cls.info,
        )
    )
    cls.defs.body[dec_index] = new_stmt


def _scan_declarative_assignment_stmt(
    cls: ClassDef,
    api: SemanticAnalyzerPluginInterface,
    stmt: AssignmentStmt,
    attributes: List[util.SQLAlchemyAttribute],
) -> None:
    """Extract mapping information from an assignment statement in a
    declarative class.

    """
    lvalue = stmt.lvalues[0]
    if not isinstance(lvalue, NameExpr):
        return

    sym = cls.info.names.get(lvalue.name)

    # this establishes that semantic analysis has taken place, which
    # means the nodes are populated and we are called from an appropriate
    # hook.
    assert sym is not None
    node = sym.node

    if isinstance(node, PlaceholderNode):
        return

    assert node is lvalue.node
    assert isinstance(node, Var)

    if node.name == "__abstract__":
        if api.parse_bool(stmt.rvalue) is True:
            util.set_is_base(cls.info)
        return
    elif node.name == "__tablename__":
        util.set_has_table(cls.info)
    elif node.name.startswith("__"):
        return
    elif node.name == "_mypy_mapped_attrs":
        if not isinstance(stmt.rvalue, ListExpr):
            util.fail(api, "_mypy_mapped_attrs is expected to be a list", stmt)
        else:
            for item in stmt.rvalue.items:
                if isinstance(item, (NameExpr, StrExpr)):
                    apply.apply_mypy_mapped_attr(cls, api, item, attributes)

    left_hand_mapped_type: Optional[Type] = None
    left_hand_explicit_type: Optional[ProperType] = None

    if node.is_inferred or node.type is None:
        if isinstance(stmt.type, UnboundType):
            # look for an explicit Mapped[] type annotation on the left
            # side with nothing on the right

            # print(stmt.type)
            # Mapped?[Optional?[A?]]

            left_hand_explicit_type = stmt.type

            if stmt.type.name == "Mapped":
                mapped_sym = api.lookup_qualified("Mapped", cls)
                if (
                    mapped_sym is not None
                    and mapped_sym.node is not None
                    and names.type_id_for_named_node(mapped_sym.node)
                    is names.MAPPED
                ):
                    left_hand_explicit_type = get_proper_type(
                        stmt.type.args[0]
                    )
                    left_hand_mapped_type = stmt.type

            # TODO: do we need to convert from unbound for this case?
            # left_hand_explicit_type = util._unbound_to_instance(
            #     api, left_hand_explicit_type
            # )
    else:
        node_type = get_proper_type(node.type)
        if (
            isinstance(node_type, Instance)
            and names.type_id_for_named_node(node_type.type) is names.MAPPED
        ):
            # print(node.type)
            # sqlalchemy.orm.attributes.Mapped[<python type>]
            left_hand_explicit_type = get_proper_type(node_type.args[0])
            left_hand_mapped_type = node_type
        else:
            # print(node.type)
            # <python type>
            left_hand_explicit_type = node_type
            left_hand_mapped_type = None

    if isinstance(stmt.rvalue, TempNode) and left_hand_mapped_type is not None:
        # annotation without assignment and Mapped is present
        # as type annotation
        # equivalent to using _infer_type_from_left_hand_type_only.

        python_type_for_type = left_hand_explicit_type
    elif isinstance(stmt.rvalue, CallExpr) and isinstance(
        stmt.rvalue.callee, RefExpr
    ):
        python_type_for_type = infer.infer_type_from_right_hand_nameexpr(
            api, stmt, node, left_hand_explicit_type, stmt.rvalue.callee
        )

        if python_type_for_type is None:
            return

    else:
        return

    assert python_type_for_type is not None

    attributes.append(
        util.SQLAlchemyAttribute(
            name=node.name,
            line=stmt.line,
            column=stmt.column,
            typ=python_type_for_type,
            info=cls.info,
        )
    )

    apply.apply_type_to_mapped_statement(
        api,
        stmt,
        lvalue,
        left_hand_explicit_type,
        python_type_for_type,
    )


def _scan_for_mapped_bases(
    cls: ClassDef,
    api: SemanticAnalyzerPluginInterface,
) -> None:
    """Given a class, iterate through its superclass hierarchy to find
    all other classes that are considered as ORM-significant.

    Locates non-mapped mixins and scans them for mapped attributes to be
    applied to subclasses.

    """

    info = util.info_for_cls(cls, api)

    if info is None:
        return

    for base_info in info.mro[1:-1]:
        if base_info.fullname.startswith("builtins"):
            continue

        # scan each base for mapped attributes.  if they are not already
        # scanned (but have all their type info), that means they are unmapped
        # mixins
        scan_declarative_assignments_and_apply_types(
            base_info.defn, api, is_mixin_scan=True
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\options.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

import warnings
from abc import ABCMeta, abstractmethod
from enum import Enum
from typing import List, Optional

from selenium.common.exceptions import InvalidArgumentException
from selenium.webdriver.common.proxy import Proxy


class PageLoadStrategy(str, Enum):
    """Enum of possible page load strategies.

    Selenium support following strategies:
        * normal (default) - waits for all resources to download
        * eager - DOM access is ready, but other resources like images may still be loading
        * none - does not block `WebDriver` at all

    Docs: https://www.selenium.dev/documentation/webdriver/drivers/options/#pageloadstrategy.
    """

    normal = "normal"
    eager = "eager"
    none = "none"


class _BaseOptionsDescriptor:
    def __init__(self, name):
        self.name = name

    def __get__(self, obj, cls):
        if self.name == "enableBidi":
            # whether BiDi is or will be enabled
            value = obj._caps.get("webSocketUrl")
            return value is True or isinstance(value, str)
        if self.name == "webSocketUrl":
            # Return socket url or None if not created yet
            value = obj._caps.get(self.name)
            return None if not isinstance(value, str) else value
        if self.name in ("acceptInsecureCerts", "strictFileInteractability", "setWindowRect", "se:downloadsEnabled"):
            return obj._caps.get(self.name, False)
        return obj._caps.get(self.name)

    def __set__(self, obj, value):
        if self.name == "enableBidi":
            obj.set_capability("webSocketUrl", value)
        else:
            obj.set_capability(self.name, value)


class _PageLoadStrategyDescriptor:
    """Determines the point at which a navigation command is returned:
    https://w3c.github.io/webdriver/#dfn-table-of-page-load-strategies.

    :param strategy: the strategy corresponding to a document readiness state
    """

    def __init__(self, name):
        self.name = name

    def __get__(self, obj, cls):
        return obj._caps.get(self.name)

    def __set__(self, obj, value):
        if value in ("normal", "eager", "none"):
            obj.set_capability(self.name, value)
        else:
            raise ValueError("Strategy can only be one of the following: normal, eager, none")


class _UnHandledPromptBehaviorDescriptor:
    """How the driver should respond when an alert is present and the:
    command sent is not handling the alert:
    https://w3c.github.io/webdriver/#dfn-table-of-page-load-strategies:

    :param behavior: behavior to use when an alert is encountered

    :returns: Values for implicit timeout, pageLoad timeout and script timeout if set (in milliseconds)
    """

    def __init__(self, name):
        self.name = name

    def __get__(self, obj, cls):
        return obj._caps.get(self.name)

    def __set__(self, obj, value):
        if value in ("dismiss", "accept", "dismiss and notify", "accept and notify", "ignore"):
            obj.set_capability(self.name, value)
        else:
            raise ValueError(
                "Behavior can only be one of the following: dismiss, accept, dismiss and notify, "
                "accept and notify, ignore"
            )


class _TimeoutsDescriptor:
    """How long the driver should wait for actions to complete before:
    returning an error https://w3c.github.io/webdriver/#timeouts:

    :param timeouts: values in milliseconds for implicit wait, page load and script timeout

    :returns: Values for implicit timeout, pageLoad timeout and script timeout if set (in milliseconds)
    """

    def __init__(self, name):
        self.name = name

    def __get__(self, obj, cls):
        return obj._caps.get(self.name)

    def __set__(self, obj, value):
        if all(x in ("implicit", "pageLoad", "script") for x in value.keys()):
            obj.set_capability(self.name, value)
        else:
            raise ValueError("Timeout keys can only be one of the following: implicit, pageLoad, script")


class _ProxyDescriptor:
    """:Returns: Proxy if set, otherwise None."""

    def __init__(self, name):
        self.name = name

    def __get__(self, obj, cls):
        return obj._proxy

    def __set__(self, obj, value):
        if not isinstance(value, Proxy):
            raise InvalidArgumentException("Only Proxy objects can be passed in.")
        obj._proxy = value
        obj._caps[self.name] = value.to_capabilities()


class BaseOptions(metaclass=ABCMeta):
    """Base class for individual browser options."""

    browser_version = _BaseOptionsDescriptor("browserVersion")
    """Gets and Sets the version of the browser.

    Usage:
    ------
    - Get
        - `self.browser_version`
    - Set
        - `self.browser_version` = `value`

    Parameters:
    -----------
    `value`: `str`

    Returns:
    --------
    - Get
        - `str`
    - Set
        - `None`
    """

    platform_name = _BaseOptionsDescriptor("platformName")
    """Gets and Sets name of the platform.

    Usage:
    ------
    - Get
        - `self.platform_name`
    - Set
        - `self.platform_name` = `value`

    Parameters:
    -----------
    `value`: `str`

    Returns:
    --------
    - Get
        - `str`
    - Set
        - `None`
    """

    accept_insecure_certs = _BaseOptionsDescriptor("acceptInsecureCerts")
    """Gets and Set whether the session accepts insecure certificates.

    Usage:
    ------
    - Get
        - `self.accept_insecure_certs`
    - Set
        - `self.accept_insecure_certs` = `value`

    Parameters:
    -----------
    `value`: `bool`

    Returns:
    --------
    - Get
        - `bool`
    - Set
        - `None`
    """

    strict_file_interactability = _BaseOptionsDescriptor("strictFileInteractability")
    """Gets and Sets whether session is about file interactability.

    Usage:
    ------
    - Get
        - `self.strict_file_interactability`
    - Set
        - `self.strict_file_interactability` = `value`

    Parameters:
    -----------
    `value`: `bool`

    Returns:
    --------
    - Get
        - `bool`
    - Set
        - `None`
    """

    set_window_rect = _BaseOptionsDescriptor("setWindowRect")
    """Gets and Sets window size and position.

    Usage:
    ------
    - Get
        - `self.set_window_rect`
    - Set
        - `self.set_window_rect` = `value`

    Parameters:
    -----------
    `value`: `bool`

    Returns:
    --------
    - Get
        - `bool`
    - Set
        - `None`
    """

    enable_bidi = _BaseOptionsDescriptor("enableBidi")
    """Gets and Set whether the session has WebDriverBiDi enabled.

    Usage:
    ------
    - Get
        - `self.enable_bidi`
    - Set
        - `self.enable_bidi` = `value`

    Parameters:
    -----------
    `value`: `bool`

    Returns:
    --------
    - Get
        - `bool`
    - Set
        - `None`
    """

    page_load_strategy = _PageLoadStrategyDescriptor("pageLoadStrategy")
    """:Gets and Sets page load strategy, the default is "normal".

    Usage:
    ------
    - Get
        - `self.page_load_strategy`
    - Set
        - `self.page_load_strategy` = `value`

    Parameters:
    -----------
    `value`: `str`

    Returns:
    --------
    - Get
        - `str`
    - Set
        - `None`
    """

    unhandled_prompt_behavior = _UnHandledPromptBehaviorDescriptor("unhandledPromptBehavior")
    """:Gets and Sets unhandled prompt behavior, the default is "dismiss and
    notify".

    Usage:
    ------
    - Get
        - `self.unhandled_prompt_behavior`
    - Set
        - `self.unhandled_prompt_behavior` = `value`

    Parameters:
    -----------
    `value`: `str`

    Returns:
    --------
    - Get
        - `str`
    - Set
        - `None`
    """

    timeouts = _TimeoutsDescriptor("timeouts")
    """:Gets and Sets implicit timeout, pageLoad timeout and script timeout if
    set (in milliseconds)

    Usage:
    ------
    - Get
        - `self.timeouts`
    - Set
        - `self.timeouts` = `value`

    Parameters:
    -----------
    `value`: `dict`

    Returns:
    --------
    - Get
        - `dict`
    - Set
        - `None`
    """

    proxy = _ProxyDescriptor("proxy")
    """Sets and Gets Proxy.

    Usage:
    ------
    - Get
        - `self.proxy`
    - Set
        - `self.proxy` = `value`

    Parameters:
    -----------
    `value`: `Proxy`

    Returns:
    --------
    - Get
        - `Proxy`
    - Set
        - `None`
    """

    enable_downloads = _BaseOptionsDescriptor("se:downloadsEnabled")
    """Gets and Sets whether session can download files.

    Usage:
    ------
    - Get
        - `self.enable_downloads`
    - Set
        - `self.enable_downloads` = `value`

    Parameters:
    -----------
    `value`: `bool`

    Returns:
    --------
    - Get
        - `bool`
    - Set
        - `None`
    """

    web_socket_url = _BaseOptionsDescriptor("webSocketUrl")
    """Gets and Sets WebSocket URL.

    Usage:
    ------
    - Get
        - `self.web_socket_url`
    - Set
        - `self.web_socket_url` = `value`

    Parameters:
    -----------
    `value`: `str`

    Returns:
    --------
    - Get
        - `bool`
    - Set
        - `None`
    """

    def __init__(self) -> None:
        super().__init__()
        self._caps = self.default_capabilities
        self._proxy = None
        self.set_capability("pageLoadStrategy", PageLoadStrategy.normal)
        self.mobile_options = None
        self._ignore_local_proxy = False

    @property
    def capabilities(self):
        return self._caps

    def set_capability(self, name, value) -> None:
        """Sets a capability."""
        self._caps[name] = value

    def enable_mobile(
        self,
        android_package: Optional[str] = None,
        android_activity: Optional[str] = None,
        device_serial: Optional[str] = None,
    ) -> None:
        """Enables mobile browser use for browsers that support it.

        :Args:
            android_activity: The name of the android package to start
        """
        if not android_package:
            raise AttributeError("android_package must be passed in")
        self.mobile_options = {"androidPackage": android_package}
        if android_activity:
            self.mobile_options["androidActivity"] = android_activity
        if device_serial:
            self.mobile_options["androidDeviceSerial"] = device_serial

    @abstractmethod
    def to_capabilities(self):
        """Convert options into capabilities dictionary."""

    @property
    @abstractmethod
    def default_capabilities(self):
        """Return minimal capabilities necessary as a dictionary."""

    def ignore_local_proxy_environment_variables(self) -> None:
        """By calling this you will ignore HTTP_PROXY and HTTPS_PROXY from
        being picked up and used."""
        self._ignore_local_proxy = True


class ArgOptions(BaseOptions):
    BINARY_LOCATION_ERROR = "Binary Location Must be a String"
    # FedCM capability key
    FEDCM_CAPABILITY = "fedcm:accounts"

    def __init__(self) -> None:
        super().__init__()
        self._arguments: List[str] = []

    @property
    def arguments(self):
        """:Returns: A list of arguments needed for the browser."""
        return self._arguments

    def add_argument(self, argument: str) -> None:
        """Adds an argument to the list.

        :Args:
         - Sets the arguments
        """
        if argument:
            self._arguments.append(argument)
        else:
            raise ValueError("argument can not be null")

    def ignore_local_proxy_environment_variables(self) -> None:
        """By calling this you will ignore HTTP_PROXY and HTTPS_PROXY from
        being picked up and used."""
        warnings.warn(
            "using ignore_local_proxy_environment_variables in Options has been deprecated, "
            "instead, create a Proxy instance with ProxyType.DIRECT to ignore proxy settings, "
            "pass the proxy instance into a ClientConfig constructor, "
            "pass the client config instance into the Webdriver constructor",
            DeprecationWarning,
            stacklevel=2,
        )

        super().ignore_local_proxy_environment_variables()

    def to_capabilities(self):
        return self._caps

    @property
    def default_capabilities(self):
        return {}

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\bidi\cdp.py ===
# The MIT License(MIT)
#
# Copyright(c) 2018 Hyperion Gray
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files(the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and / or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
#
# This code comes from https://github.com/HyperionGray/trio-chrome-devtools-protocol/tree/master/trio_cdp

# flake8: noqa

import contextvars
import importlib
import itertools
import json
import logging
import pathlib
from collections import defaultdict
from contextlib import asynccontextmanager
from contextlib import contextmanager
from dataclasses import dataclass
from typing import Any
from typing import AsyncGenerator
from typing import AsyncIterator
from typing import Generator
from typing import Type
from typing import TypeVar

import trio
from trio_websocket import ConnectionClosed as WsConnectionClosed
from trio_websocket import connect_websocket_url

logger = logging.getLogger("trio_cdp")
T = TypeVar("T")
MAX_WS_MESSAGE_SIZE = 2**24

devtools = None
version = None


def import_devtools(ver):
    """Attempt to load the current latest available devtools into the module
    cache for use later."""
    global devtools
    global version
    version = ver
    base = "selenium.webdriver.common.devtools.v"
    try:
        devtools = importlib.import_module(f"{base}{ver}")
        return devtools
    except ModuleNotFoundError:
        # Attempt to parse and load the 'most recent' devtools module. This is likely
        # because cdp has been updated but selenium python has not been released yet.
        devtools_path = pathlib.Path(__file__).parents[1].joinpath("devtools")
        versions = tuple(f.name for f in devtools_path.iterdir() if f.is_dir())
        latest = max(int(x[1:]) for x in versions)
        selenium_logger = logging.getLogger(__name__)
        selenium_logger.debug("Falling back to loading `devtools`: v%s", latest)
        devtools = importlib.import_module(f"{base}{latest}")
        return devtools


_connection_context: contextvars.ContextVar = contextvars.ContextVar("connection_context")
_session_context: contextvars.ContextVar = contextvars.ContextVar("session_context")


def get_connection_context(fn_name):
    """Look up the current connection.

    If there is no current connection, raise a ``RuntimeError`` with a
    helpful message.
    """
    try:
        return _connection_context.get()
    except LookupError:
        raise RuntimeError(f"{fn_name}() must be called in a connection context.")


def get_session_context(fn_name):
    """Look up the current session.

    If there is no current session, raise a ``RuntimeError`` with a
    helpful message.
    """
    try:
        return _session_context.get()
    except LookupError:
        raise RuntimeError(f"{fn_name}() must be called in a session context.")


@contextmanager
def connection_context(connection):
    """This context manager installs ``connection`` as the session context for
    the current Trio task."""
    token = _connection_context.set(connection)
    try:
        yield
    finally:
        _connection_context.reset(token)


@contextmanager
def session_context(session):
    """This context manager installs ``session`` as the session context for the
    current Trio task."""
    token = _session_context.set(session)
    try:
        yield
    finally:
        _session_context.reset(token)


def set_global_connection(connection):
    """Install ``connection`` in the root context so that it will become the
    default connection for all tasks.

    This is generally not recommended, except it may be necessary in
    certain use cases such as running inside Jupyter notebook.
    """
    global _connection_context
    _connection_context = contextvars.ContextVar("_connection_context", default=connection)


def set_global_session(session):
    """Install ``session`` in the root context so that it will become the
    default session for all tasks.

    This is generally not recommended, except it may be necessary in
    certain use cases such as running inside Jupyter notebook.
    """
    global _session_context
    _session_context = contextvars.ContextVar("_session_context", default=session)


class BrowserError(Exception):
    """This exception is raised when the browser's response to a command
    indicates that an error occurred."""

    def __init__(self, obj):
        self.code = obj.get("code")
        self.message = obj.get("message")
        self.detail = obj.get("data")

    def __str__(self):
        return f"BrowserError<code={self.code} message={self.message}> {self.detail}"


class CdpConnectionClosed(WsConnectionClosed):
    """Raised when a public method is called on a closed CDP connection."""

    def __init__(self, reason):
        """Constructor.

        :param reason:
        :type reason: wsproto.frame_protocol.CloseReason
        """
        self.reason = reason

    def __repr__(self):
        """Return representation."""
        return f"{self.__class__.__name__}<{self.reason}>"


class InternalError(Exception):
    """This exception is only raised when there is faulty logic in TrioCDP or
    the integration with PyCDP."""


@dataclass
class CmEventProxy:
    """A proxy object returned by :meth:`CdpBase.wait_for()``.

    After the context manager executes, this proxy object will have a
    value set that contains the returned event.
    """

    value: Any = None


class CdpBase:
    def __init__(self, ws, session_id, target_id):
        self.ws = ws
        self.session_id = session_id
        self.target_id = target_id
        self.channels = defaultdict(set)
        self.id_iter = itertools.count()
        self.inflight_cmd = {}
        self.inflight_result = {}

    async def execute(self, cmd: Generator[dict, T, Any]) -> T:
        """Execute a command on the server and wait for the result.

        :param cmd: any CDP command
        :returns: a CDP result
        """
        cmd_id = next(self.id_iter)
        cmd_event = trio.Event()
        self.inflight_cmd[cmd_id] = cmd, cmd_event
        request = next(cmd)
        request["id"] = cmd_id
        if self.session_id:
            request["sessionId"] = self.session_id
        request_str = json.dumps(request)
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug(f"Sending CDP message: {cmd_id} {cmd_event}: {request_str}")
        try:
            await self.ws.send_message(request_str)
        except WsConnectionClosed as wcc:
            raise CdpConnectionClosed(wcc.reason) from None
        await cmd_event.wait()
        response = self.inflight_result.pop(cmd_id)
        if logger.isEnabledFor(logging.DEBUG):
            logger.debug(f"Received CDP message: {response}")
        if isinstance(response, Exception):
            if logger.isEnabledFor(logging.DEBUG):
                logger.debug(f"Exception raised by {cmd_event} message: {type(response).__name__}")
            raise response
        return response

    def listen(self, *event_types, buffer_size=10):
        """Return an async iterator that iterates over events matching the
        indicated types."""
        sender, receiver = trio.open_memory_channel(buffer_size)
        for event_type in event_types:
            self.channels[event_type].add(sender)
        return receiver

    @asynccontextmanager
    async def wait_for(self, event_type: Type[T], buffer_size=10) -> AsyncGenerator[CmEventProxy, None]:
        """Wait for an event of the given type and return it.

        This is an async context manager, so you should open it inside
        an async with block. The block will not exit until the indicated
        event is received.
        """
        sender: trio.MemorySendChannel
        receiver: trio.MemoryReceiveChannel
        sender, receiver = trio.open_memory_channel(buffer_size)
        self.channels[event_type].add(sender)
        proxy = CmEventProxy()
        yield proxy
        async with receiver:
            event = await receiver.receive()
        proxy.value = event

    def _handle_data(self, data):
        """Handle incoming WebSocket data.

        :param dict data: a JSON dictionary
        """
        if "id" in data:
            self._handle_cmd_response(data)
        else:
            self._handle_event(data)

    def _handle_cmd_response(self, data):
        """Handle a response to a command. This will set an event flag that
        will return control to the task that called the command.

        :param dict data: response as a JSON dictionary
        """
        cmd_id = data["id"]
        try:
            cmd, event = self.inflight_cmd.pop(cmd_id)
        except KeyError:
            logger.warning("Got a message with a command ID that does not exist: %s", data)
            return
        if "error" in data:
            # If the server reported an error, convert it to an exception and do
            # not process the response any further.
            self.inflight_result[cmd_id] = BrowserError(data["error"])
        else:
            # Otherwise, continue the generator to parse the JSON result
            # into a CDP object.
            try:
                _ = cmd.send(data["result"])
                raise InternalError("The command's generator function did not exit when expected!")
            except StopIteration as exit:
                return_ = exit.value
            self.inflight_result[cmd_id] = return_
        event.set()

    def _handle_event(self, data):
        """Handle an event.

        :param dict data: event as a JSON dictionary
        """
        global devtools
        event = devtools.util.parse_json_event(data)
        logger.debug("Received event: %s", event)
        to_remove = set()
        for sender in self.channels[type(event)]:
            try:
                sender.send_nowait(event)
            except trio.WouldBlock:
                logger.error('Unable to send event "%r" due to full channel %s', event, sender)
            except trio.BrokenResourceError:
                to_remove.add(sender)
        if to_remove:
            self.channels[type(event)] -= to_remove


class CdpSession(CdpBase):
    """Contains the state for a CDP session.

    Generally you should not instantiate this object yourself; you should call
    :meth:`CdpConnection.open_session`.
    """

    def __init__(self, ws, session_id, target_id):
        """Constructor.

        :param trio_websocket.WebSocketConnection ws:
        :param devtools.target.SessionID session_id:
        :param devtools.target.TargetID target_id:
        """
        super().__init__(ws, session_id, target_id)

        self._dom_enable_count = 0
        self._dom_enable_lock = trio.Lock()
        self._page_enable_count = 0
        self._page_enable_lock = trio.Lock()

    @asynccontextmanager
    async def dom_enable(self):
        """A context manager that executes ``dom.enable()`` when it enters and
        then calls ``dom.disable()``.

        This keeps track of concurrent callers and only disables DOM
        events when all callers have exited.
        """
        global devtools
        async with self._dom_enable_lock:
            self._dom_enable_count += 1
            if self._dom_enable_count == 1:
                await self.execute(devtools.dom.enable())

        yield

        async with self._dom_enable_lock:
            self._dom_enable_count -= 1
            if self._dom_enable_count == 0:
                await self.execute(devtools.dom.disable())

    @asynccontextmanager
    async def page_enable(self):
        """A context manager that executes ``page.enable()`` when it enters and
        then calls ``page.disable()`` when it exits.

        This keeps track of concurrent callers and only disables page
        events when all callers have exited.
        """
        global devtools
        async with self._page_enable_lock:
            self._page_enable_count += 1
            if self._page_enable_count == 1:
                await self.execute(devtools.page.enable())

        yield

        async with self._page_enable_lock:
            self._page_enable_count -= 1
            if self._page_enable_count == 0:
                await self.execute(devtools.page.disable())


class CdpConnection(CdpBase, trio.abc.AsyncResource):
    """Contains the connection state for a Chrome DevTools Protocol server.

    CDP can multiplex multiple "sessions" over a single connection. This
    class corresponds to the "root" session, i.e. the implicitly created
    session that has no session ID. This class is responsible for
    reading incoming WebSocket messages and forwarding them to the
    corresponding session, as well as handling messages targeted at the
    root session itself. You should generally call the
    :func:`open_cdp()` instead of instantiating this class directly.
    """

    def __init__(self, ws):
        """Constructor.

        :param trio_websocket.WebSocketConnection ws:
        """
        super().__init__(ws, session_id=None, target_id=None)
        self.sessions = {}

    async def aclose(self):
        """Close the underlying WebSocket connection.

        This will cause the reader task to gracefully exit when it tries
        to read the next message from the WebSocket. All of the public
        APIs (``execute()``, ``listen()``, etc.) will raise
        ``CdpConnectionClosed`` after the CDP connection is closed. It
        is safe to call this multiple times.
        """
        await self.ws.aclose()

    @asynccontextmanager
    async def open_session(self, target_id) -> AsyncIterator[CdpSession]:
        """This context manager opens a session and enables the "simple" style
        of calling CDP APIs.

        For example, inside a session context, you can call ``await
        dom.get_document()`` and it will execute on the current session
        automatically.
        """
        session = await self.connect_session(target_id)
        with session_context(session):
            yield session

    async def connect_session(self, target_id) -> "CdpSession":
        """Returns a new :class:`CdpSession` connected to the specified
        target."""
        global devtools
        session_id = await self.execute(devtools.target.attach_to_target(target_id, True))
        session = CdpSession(self.ws, session_id, target_id)
        self.sessions[session_id] = session
        return session

    async def _reader_task(self):
        """Runs in the background and handles incoming messages: dispatching
        responses to commands and events to listeners."""
        global devtools
        while True:
            try:
                message = await self.ws.get_message()
            except WsConnectionClosed:
                # If the WebSocket is closed, we don't want to throw an
                # exception from the reader task. Instead we will throw
                # exceptions from the public API methods, and we can quietly
                # exit the reader task here.
                break
            try:
                data = json.loads(message)
            except json.JSONDecodeError:
                raise BrowserError({"code": -32700, "message": "Client received invalid JSON", "data": message})
            logger.debug("Received message %r", data)
            if "sessionId" in data:
                session_id = devtools.target.SessionID(data["sessionId"])
                try:
                    session = self.sessions[session_id]
                except KeyError:
                    raise BrowserError(
                        {
                            "code": -32700,
                            "message": "Browser sent a message for an invalid session",
                            "data": f"{session_id!r}",
                        }
                    )
                session._handle_data(data)
            else:
                self._handle_data(data)

        for _, session in self.sessions.items():
            for _, senders in session.channels.items():
                for sender in senders:
                    sender.close()


@asynccontextmanager
async def open_cdp(url) -> AsyncIterator[CdpConnection]:
    """This async context manager opens a connection to the browser specified
    by ``url`` before entering the block, then closes the connection when the
    block exits.

    The context manager also sets the connection as the default
    connection for the current task, so that commands like ``await
    target.get_targets()`` will run on this connection automatically. If
    you want to use multiple connections concurrently, it is recommended
    to open each on in a separate task.
    """

    async with trio.open_nursery() as nursery:
        conn = await connect_cdp(nursery, url)
        try:
            with connection_context(conn):
                yield conn
        finally:
            await conn.aclose()


async def connect_cdp(nursery, url) -> CdpConnection:
    """Connect to the browser specified by ``url`` and spawn a background task
    in the specified nursery.

    The ``open_cdp()`` context manager is preferred in most situations.
    You should only use this function if you need to specify a custom
    nursery. This connection is not automatically closed! You can either
    use the connection object as a context manager (``async with
    conn:``) or else call ``await conn.aclose()`` on it when you are
    done with it. If ``set_context`` is True, then the returned
    connection will be installed as the default connection for the
    current task. This argument is for unusual use cases, such as
    running inside of a notebook.
    """
    ws = await connect_websocket_url(nursery, url, max_message_size=MAX_WS_MESSAGE_SIZE)
    cdp_conn = CdpConnection(ws)
    nursery.start_soon(cdp_conn._reader_task)
    return cdp_conn

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\gpt2\gpt2_parity.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.  See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------

# This script uses different configurations in mixed precision conversion for GPT-2 model, and
# measures the inference latency, top 1 match rate (compared to PyTorch FP32 model) and ONNX model size.
# It outputs a csv file with Mann-Whitney U test and T-Test on each pair of experiments, where
# pvalue < 0.05 means two experiments have significant difference on top 1 match rate.
# User could use this script to select the best mixed precision model according to these metrics.

import argparse
import csv
import datetime
import json
import logging
import os

import onnx
import scipy.stats
from benchmark_helper import get_ort_environment_variables, setup_logger
from convert_to_onnx import main
from gpt2_helper import PRETRAINED_GPT2_MODELS, Gpt2Helper
from onnx_model import OnnxModel

logger = logging.getLogger("")


def parse_arguments(argv=None):
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "-m",
        "--model_name_or_path",
        required=True,
        type=str,
        help="Model path, or pretrained model name in the list: " + ", ".join(PRETRAINED_GPT2_MODELS),
    )

    parser.add_argument(
        "--csv",
        required=False,
        type=str,
        default="gpt2_parity_results.csv",
        help="path of csv file to save the result",
    )

    parser.add_argument(
        "--test_cases",
        required=False,
        type=int,
        default=500,
        help="number of test cases per run",
    )

    parser.add_argument("--runs", required=False, type=int, default=40, help="number of repeated runs")

    parser.add_argument("--use_gpu", required=False, action="store_true", help="use GPU for inference")
    parser.set_defaults(use_gpu=False)

    parser.add_argument(
        "--all",
        required=False,
        action="store_true",
        help="run all combinations of mixed precision",
    )
    parser.set_defaults(all=False)

    parser.add_argument("-e", "--use_external_data_format", required=False, action="store_true")
    parser.set_defaults(use_external_data_format=False)

    parser.add_argument("--verbose", required=False, action="store_true")
    parser.set_defaults(verbose=False)

    parser.add_argument(
        "--skip_test",
        required=False,
        action="store_true",
        help="do not run test, and only rank experiments based on existing csv file",
    )
    parser.set_defaults(skip_test=False)

    parser.add_argument(
        "--overwrite",
        required=False,
        action="store_true",
        help="Overwrite existing csv file",
    )
    parser.set_defaults(overwrite=False)

    args = parser.parse_args(argv)

    return args


class ParityTask:
    def __init__(self, test_cases, total_runs, csv_path):
        self.total_runs = total_runs
        self.test_cases = test_cases
        self.csv_path = csv_path
        self.results = []
        self.run_id = 0

    def run(self, argv, experiment_name):
        start_time = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        run_id = f"{start_time}_{self.run_id}"
        self.run_id += 1

        try:
            result = main(
                [*argv, "-t", f"{self.test_cases}", "-r", f"{self.total_runs}"],
                experiment_name=experiment_name,
                run_id=run_id,
                csv_filename=self.csv_path,
            )
            if result:
                self.results.append(result)
        except Exception:
            logger.exception(f"Failed to run experiment {experiment_name}")
            result = None

        return result


def load_results_from_csv(csv_path):
    rows = []
    import csv

    with open(csv_path, newline="") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            rows.append(row)  # noqa: PERF402
    return rows


def get_latency(row):
    for name in row:
        if name.startswith("average_latency(batch_size="):
            return float(row[name])

    raise RuntimeError("Failed to get average_latency from output")


def score(row):
    """Scoring function based on 3 metrics. The larger score is better."""
    latency_in_ms = get_latency(row)
    top1_match_rate = float(row["top1_match_rate"])
    onnx_size_in_MB = float(row["onnx_size_in_MB"])  # noqa: N806
    # A simple scoring function: cost of 0.1ms latency ~ 0.1% match rate ~ 100MB size
    return top1_match_rate * 1000 - latency_in_ms * 10 - onnx_size_in_MB / 100


def print_wins(wins, rows, test_name):
    print()
    print("*" * 10)

    row_map = {}
    for row in rows:
        row_map[row["run_id"]] = row

    sorted_wins = dict(
        sorted(
            wins.items(),
            key=lambda item: (item[1], score(row_map[item[0]])),
            reverse=True,
        )
    )
    logger.debug(f"{test_name} Wins:{sorted_wins}")
    logger.info(f"Based on {test_name} wins and a scoring function, the ranking:")

    rank = 0
    previous_value = -1
    for count, (key, value) in enumerate(sorted_wins.items()):
        if value != previous_value:
            rank = count
        previous_value = value

        for row in rows:
            if row["run_id"] == key:
                logger.info(
                    "{:02d}: WINs={:02d}, run_id={}, latency={:5.2f}, top1_match={:.4f}, size={}_MB, experiment={}, {}".format(  # noqa: G001
                        rank,
                        value,
                        key,
                        get_latency(row),
                        float(row["top1_match_rate"]),
                        row["onnx_size_in_MB"],
                        row["experiment"],
                        get_ort_environment_variables(),
                    )
                )
                break


def run_significance_test(rows, output_csv_path):
    """Run U test and T test."""
    utest_wins = {}
    ttest_wins = {}
    for row in rows:
        run_id = row["run_id"]
        utest_wins[run_id] = 0
        ttest_wins[run_id] = 0

    with open(output_csv_path, "w", newline="") as csvfile:
        column_names = [
            "model_name",
            "run_id_1",
            "experiment_1",
            "top1_match_rate_1",
            "run_id_2",
            "experiment_2",
            "top1_match_rate_2",
            "U_statistic",
            "U_pvalue",
            "T_statistic",
            "T_pvalue",
        ]

        writer = csv.DictWriter(csvfile, fieldnames=column_names)
        writer.writeheader()

        required_match_columns = ["model_name", "test_cases", "runs"]
        num_results = len(rows)
        for i in range(num_results - 1):
            result1 = rows[i]

            if isinstance(result1["top1_match_rate_per_run"], str):
                a = json.loads(result1["top1_match_rate_per_run"])
            else:
                a = result1["top1_match_rate_per_run"]

            for j in range(i + 1, num_results, 1):
                result2 = rows[j]

                all_matched = True
                for column in required_match_columns:
                    if result1[column] != result2[column]:
                        all_matched = False
                        break
                if not all_matched:
                    continue

                if isinstance(result2["top1_match_rate_per_run"], str):
                    b = json.loads(result2["top1_match_rate_per_run"])
                else:
                    b = result2["top1_match_rate_per_run"]

                try:
                    utest_statistic, utest_pvalue = scipy.stats.mannwhitneyu(
                        a, b, use_continuity=True, alternative="two-sided"
                    )  # TODO: shall we use one-sided: less or greater according to "top1_match_rate"
                except ValueError:  # ValueError: All numbers are identical in mannwhitneyu
                    utest_statistic = None
                    utest_pvalue = None
                ttest_statistic, ttest_pvalue = scipy.stats.ttest_ind(a, b, axis=None, equal_var=True)

                if utest_pvalue is not None and utest_pvalue < 0.05:
                    if float(result1["top1_match_rate"]) > float(result2["top1_match_rate"]):
                        utest_wins[result1["run_id"]] += 1
                    else:
                        utest_wins[result2["run_id"]] += 1

                if ttest_pvalue < 0.05:
                    if float(result1["top1_match_rate"]) > float(result2["top1_match_rate"]):
                        ttest_wins[result1["run_id"]] += 1
                    else:
                        ttest_wins[result2["run_id"]] += 1

                row = {
                    "model_name": result1["model_name"],
                    "run_id_1": result1["run_id"],
                    "experiment_1": result1["experiment"],
                    "top1_match_rate_1": float(result1["top1_match_rate"]),
                    "run_id_2": result2["run_id"],
                    "experiment_2": result2["experiment"],
                    "top1_match_rate_2": float(result2["top1_match_rate"]),
                    "U_statistic": utest_statistic,
                    "U_pvalue": utest_pvalue,
                    "T_statistic": ttest_statistic,
                    "T_pvalue": ttest_pvalue,
                }

                writer.writerow(row)
    logger.info(f"U-Test and T-Test results are output to {output_csv_path}")
    print_wins(utest_wins, rows, "U-Test")
    print_wins(ttest_wins, rows, "T-Test")


def get_last_matmul_node_name(raw_onnx_model: str):
    model = onnx.load(raw_onnx_model)
    onnx_model = OnnxModel(model)
    output_name_to_node = onnx_model.output_name_to_node()

    assert model.graph.output[0].name in output_name_to_node
    node = output_name_to_node[model.graph.output[0].name]
    if node.op_type == "MatMul":
        logger.info(f"Found last MatMul node for logits: {node.name}")
        return node.name

    logger.warning(f"Failed to find MatMul node for logits. Found {node.op_type} of node {node.name}")
    return None


def get_mixed_precision_parameters(args, last_matmul_node_name, op_block_list):
    model = args.model_name_or_path
    parameters = f"-m {model} -o --use_gpu -p fp16".split()
    if args.use_external_data_format:
        parameters.append("--use_external_data_format")
    parameters += [
        "--io_block_list",
        "logits",
        "--node_block_list",
        last_matmul_node_name,
    ]

    if op_block_list:
        parameters.extend(["--op_block_list", *op_block_list])

    return parameters


def run_candidate(
    task: ParityTask,
    args,
    last_matmul_node_name,
    op_block_list=["FastGelu", "LayerNormalization"],  # noqa: B006
):
    parameters = get_mixed_precision_parameters(args, last_matmul_node_name, op_block_list)
    op_block_list_str = ",".join(sorted(op_block_list))

    if op_block_list:
        name = f"Mixed precision baseline + {op_block_list_str} in FP32"
    else:
        name = f"Mixed precision baseline (logits output and last MatMul node {last_matmul_node_name} in FP32)"

    env_vars = get_ort_environment_variables()
    if env_vars:
        name = name + f" ({env_vars})"

    task.run(parameters, name)


def get_baselines(args):
    model = args.model_name_or_path
    fp32_baseline = f"-m {model} -o -p fp32".split()
    if args.use_gpu:
        fp32_baseline.append("--use_gpu")
    if args.use_external_data_format:
        fp32_baseline.append("--use_external_data_format")

    fp16_baseline = f"-m {model} -o --use_gpu -p fp16".split()
    if args.use_external_data_format:
        fp16_baseline.append("--use_external_data_format")

    return fp32_baseline, fp16_baseline


def run_tuning_step0(task, fp16_baseline, all_ops, optimized_ops):
    """Step 0 is to check which operator in FP16 causes most loss"""
    fp32_logits = ["--io_block_list", "logits"]
    task.run(fp16_baseline + fp32_logits, "FP16 except logits")

    fp32_io = ["--keep_io_types"]
    task.run(fp16_baseline + fp32_io, "Graph I/O FP32, Other FP16")

    # Only weights in FP16
    task.run(
        fp16_baseline + fp32_io + ["--op_block_list"] + list(all_ops) + ["--force_fp16_initializers"],
        "FP32 except weights in FP16",
    )

    optimized_ops_results = []
    op_list = optimized_ops
    for op in op_list:
        op_block_list = ["--op_block_list"] + [o for o in op_list if o != op]
        result = task.run(fp16_baseline + fp32_io + op_block_list, f"FP32 except {op} in FP16")
        if result:
            optimized_ops_results.append(result)

    # Check which optimized operator causes the most loss in precision
    min_result = min(optimized_ops_results, key=lambda y: y["top1_match_rate"])
    print("step 0: optimized operator causes the most loss in precision", min_result)


def run_tuning_step1(task, mixed_precision_baseline, optimized_ops):
    """Step 1 is to figure out which optimized operator in FP32 could benefit most"""
    for op in optimized_ops:
        op_block_list = ["--op_block_list", op]
        task.run(
            mixed_precision_baseline + op_block_list,
            f"Mixed precision baseline + {op} in FP32",
        )


def run_tuning_step2(task, mixed_precision_baseline, optimized_ops):
    """Assumed that you have run step 0 and 1 to figure out that Logits FP32 and some operators shall be in FP32,
    This step will try add one more operator.
    """
    candidate_fp32_ops = ["FastGelu", "LayerNormalization", "SkipLayerNormalization"]
    fp32_ops = [x for x in candidate_fp32_ops if x in optimized_ops]
    for op in optimized_ops:
        if op not in fp32_ops:
            op_block_list = [*fp32_ops, op]
            task.run(
                [*mixed_precision_baseline, "--op_block_list", *op_block_list],
                "Mixed precision baseline + {},{} in FP32".format(",".join(fp32_ops), op),
            )


def run_parity(task: ParityTask, args):
    onnx_model_paths = Gpt2Helper.get_onnx_paths(
        "onnx_models",
        args.model_name_or_path,
        new_folder=args.use_external_data_format,
        remove_existing=[],
    )

    fp32_baseline, fp16_baseline = get_baselines(args)

    result = task.run(fp32_baseline, "FP32 baseline")

    optimized_ops = []
    if result and ("optimized_operators" in result) and result["optimized_operators"]:
        optimized_ops = result["optimized_operators"].split(",")
    else:
        raise RuntimeError("Failed to get optimized operators")

    all_ops = []
    if result and ("operators" in result) and result["operators"]:
        all_ops = result["operators"].split(",")
    else:
        raise RuntimeError("Failed to get operators")

    # The following tests for fp16 requires GPU
    if not args.use_gpu:
        logger.info("skip mixed precision since --use_gpu is not specified")
        return

    task.run(fp16_baseline, "FP16 baseline")

    last_matmul_node_name = get_last_matmul_node_name(onnx_model_paths["raw"])

    # Mixed precision baseline
    run_candidate(task, args, last_matmul_node_name, op_block_list=[])

    def get_fp32_ops(x):
        return [op for op in x if op in all_ops]

    if args.all:
        run_tuning_step0(task, fp16_baseline, all_ops, optimized_ops)
        mixed_precision_baseline = get_mixed_precision_parameters(args, last_matmul_node_name, op_block_list=[])
        run_tuning_step1(task, mixed_precision_baseline, optimized_ops)
        run_tuning_step2(task, mixed_precision_baseline, optimized_ops)
    else:
        run_candidate(
            task,
            args,
            last_matmul_node_name,
            op_block_list=get_fp32_ops(["SkipLayerNormalization", "LayerNormalization", "Add"]),
        )
        run_candidate(task, args, last_matmul_node_name, op_block_list=["FastGelu"])

    # Run a few good candidates
    run_candidate(
        task,
        args,
        last_matmul_node_name,
        op_block_list=get_fp32_ops(["FastGelu", "SkipLayerNormalization", "LayerNormalization", "Add"]),
    )
    run_candidate(
        task,
        args,
        last_matmul_node_name,
        op_block_list=get_fp32_ops(
            ["FastGelu", "EmbedLayerNormalization", "SkipLayerNormalization", "LayerNormalization", "Add"]
        ),
    )


if __name__ == "__main__":
    args = parse_arguments()
    setup_logger(args.verbose)

    if args.test_cases < 100 or args.runs < 20 or args.test_cases * args.runs < 10000:
        logger.warning(
            "Not enough test cases or runs to get stable results or test significance. "
            "Recommend test_cases >= 100, runs >= 20, test_cases * runs >= 10000."
        )

    if os.path.exists(args.csv) and not args.skip_test:
        if not args.overwrite:
            raise RuntimeError(
                f"Output file {args.csv} existed. Please remove the file, or use either --skip_test or --overwrite."
            )
        else:
            logger.info("Remove existing file %s since --overwrite is specified", args.csv)
            os.remove(args.csv)

    task = ParityTask(args.test_cases, args.runs, args.csv)

    if not args.skip_test:
        run_parity(task, args)

    try:
        rows = load_results_from_csv(task.csv_path)
    except Exception:
        logger.exception(f"Failed to load csv {task.csv_path}")
        rows = task.results

    logger.info("Start running significance tests...")
    summary_csv = task.csv_path.replace(".csv", ".stats.csv")
    run_significance_test(rows, summary_csv)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\indexes\category.py ===
from __future__ import annotations

from typing import (
    TYPE_CHECKING,
    Any,
    Literal,
    cast,
)

import numpy as np

from pandas._libs import index as libindex
from pandas.util._decorators import (
    cache_readonly,
    doc,
)

from pandas.core.dtypes.common import is_scalar
from pandas.core.dtypes.concat import concat_compat
from pandas.core.dtypes.dtypes import CategoricalDtype
from pandas.core.dtypes.missing import (
    is_valid_na_for_dtype,
    isna,
)

from pandas.core.arrays.categorical import (
    Categorical,
    contains,
)
from pandas.core.construction import extract_array
from pandas.core.indexes.base import (
    Index,
    maybe_extract_name,
)
from pandas.core.indexes.extension import (
    NDArrayBackedExtensionIndex,
    inherit_names,
)

if TYPE_CHECKING:
    from collections.abc import Hashable

    from pandas._typing import (
        Dtype,
        DtypeObj,
        Self,
        npt,
    )


@inherit_names(
    [
        "argsort",
        "tolist",
        "codes",
        "categories",
        "ordered",
        "_reverse_indexer",
        "searchsorted",
        "min",
        "max",
    ],
    Categorical,
)
@inherit_names(
    [
        "rename_categories",
        "reorder_categories",
        "add_categories",
        "remove_categories",
        "remove_unused_categories",
        "set_categories",
        "as_ordered",
        "as_unordered",
    ],
    Categorical,
    wrap=True,
)
class CategoricalIndex(NDArrayBackedExtensionIndex):
    """
    Index based on an underlying :class:`Categorical`.

    CategoricalIndex, like Categorical, can only take on a limited,
    and usually fixed, number of possible values (`categories`). Also,
    like Categorical, it might have an order, but numerical operations
    (additions, divisions, ...) are not possible.

    Parameters
    ----------
    data : array-like (1-dimensional)
        The values of the categorical. If `categories` are given, values not in
        `categories` will be replaced with NaN.
    categories : index-like, optional
        The categories for the categorical. Items need to be unique.
        If the categories are not given here (and also not in `dtype`), they
        will be inferred from the `data`.
    ordered : bool, optional
        Whether or not this categorical is treated as an ordered
        categorical. If not given here or in `dtype`, the resulting
        categorical will be unordered.
    dtype : CategoricalDtype or "category", optional
        If :class:`CategoricalDtype`, cannot be used together with
        `categories` or `ordered`.
    copy : bool, default False
        Make a copy of input ndarray.
    name : object, optional
        Name to be stored in the index.

    Attributes
    ----------
    codes
    categories
    ordered

    Methods
    -------
    rename_categories
    reorder_categories
    add_categories
    remove_categories
    remove_unused_categories
    set_categories
    as_ordered
    as_unordered
    map

    Raises
    ------
    ValueError
        If the categories do not validate.
    TypeError
        If an explicit ``ordered=True`` is given but no `categories` and the
        `values` are not sortable.

    See Also
    --------
    Index : The base pandas Index type.
    Categorical : A categorical array.
    CategoricalDtype : Type for categorical data.

    Notes
    -----
    See the `user guide
    <https://pandas.pydata.org/pandas-docs/stable/user_guide/advanced.html#categoricalindex>`__
    for more.

    Examples
    --------
    >>> pd.CategoricalIndex(["a", "b", "c", "a", "b", "c"])
    CategoricalIndex(['a', 'b', 'c', 'a', 'b', 'c'],
                     categories=['a', 'b', 'c'], ordered=False, dtype='category')

    ``CategoricalIndex`` can also be instantiated from a ``Categorical``:

    >>> c = pd.Categorical(["a", "b", "c", "a", "b", "c"])
    >>> pd.CategoricalIndex(c)
    CategoricalIndex(['a', 'b', 'c', 'a', 'b', 'c'],
                     categories=['a', 'b', 'c'], ordered=False, dtype='category')

    Ordered ``CategoricalIndex`` can have a min and max value.

    >>> ci = pd.CategoricalIndex(
    ...     ["a", "b", "c", "a", "b", "c"], ordered=True, categories=["c", "b", "a"]
    ... )
    >>> ci
    CategoricalIndex(['a', 'b', 'c', 'a', 'b', 'c'],
                     categories=['c', 'b', 'a'], ordered=True, dtype='category')
    >>> ci.min()
    'c'
    """

    _typ = "categoricalindex"
    _data_cls = Categorical

    @property
    def _can_hold_strings(self):
        return self.categories._can_hold_strings

    @cache_readonly
    def _should_fallback_to_positional(self) -> bool:
        return self.categories._should_fallback_to_positional

    codes: np.ndarray
    categories: Index
    ordered: bool | None
    _data: Categorical
    _values: Categorical

    @property
    def _engine_type(self) -> type[libindex.IndexEngine]:
        # self.codes can have dtype int8, int16, int32 or int64, so we need
        # to return the corresponding engine type (libindex.Int8Engine, etc.).
        return {
            np.int8: libindex.Int8Engine,
            np.int16: libindex.Int16Engine,
            np.int32: libindex.Int32Engine,
            np.int64: libindex.Int64Engine,
        }[self.codes.dtype.type]

    # --------------------------------------------------------------------
    # Constructors

    def __new__(
        cls,
        data=None,
        categories=None,
        ordered=None,
        dtype: Dtype | None = None,
        copy: bool = False,
        name: Hashable | None = None,
    ) -> Self:
        name = maybe_extract_name(name, data, cls)

        if is_scalar(data):
            # GH#38944 include None here, which pre-2.0 subbed in []
            cls._raise_scalar_data_error(data)

        data = Categorical(
            data, categories=categories, ordered=ordered, dtype=dtype, copy=copy
        )

        return cls._simple_new(data, name=name)

    # --------------------------------------------------------------------

    def _is_dtype_compat(self, other: Index) -> Categorical:
        """
        *this is an internal non-public method*

        provide a comparison between the dtype of self and other (coercing if
        needed)

        Parameters
        ----------
        other : Index

        Returns
        -------
        Categorical

        Raises
        ------
        TypeError if the dtypes are not compatible
        """
        if isinstance(other.dtype, CategoricalDtype):
            cat = extract_array(other)
            cat = cast(Categorical, cat)
            if not cat._categories_match_up_to_permutation(self._values):
                raise TypeError(
                    "categories must match existing categories when appending"
                )

        elif other._is_multi:
            # preempt raising NotImplementedError in isna call
            raise TypeError("MultiIndex is not dtype-compatible with CategoricalIndex")
        else:
            values = other

            cat = Categorical(other, dtype=self.dtype)
            other = CategoricalIndex(cat)
            if not other.isin(values).all():
                raise TypeError(
                    "cannot append a non-category item to a CategoricalIndex"
                )
            cat = other._values

            if not ((cat == values) | (isna(cat) & isna(values))).all():
                # GH#37667 see test_equals_non_category
                raise TypeError(
                    "categories must match existing categories when appending"
                )

        return cat

    def equals(self, other: object) -> bool:
        """
        Determine if two CategoricalIndex objects contain the same elements.

        Returns
        -------
        bool
            ``True`` if two :class:`pandas.CategoricalIndex` objects have equal
            elements, ``False`` otherwise.

        Examples
        --------
        >>> ci = pd.CategoricalIndex(['a', 'b', 'c', 'a', 'b', 'c'])
        >>> ci2 = pd.CategoricalIndex(pd.Categorical(['a', 'b', 'c', 'a', 'b', 'c']))
        >>> ci.equals(ci2)
        True

        The order of elements matters.

        >>> ci3 = pd.CategoricalIndex(['c', 'b', 'a', 'a', 'b', 'c'])
        >>> ci.equals(ci3)
        False

        The orderedness also matters.

        >>> ci4 = ci.as_ordered()
        >>> ci.equals(ci4)
        False

        The categories matter, but the order of the categories matters only when
        ``ordered=True``.

        >>> ci5 = ci.set_categories(['a', 'b', 'c', 'd'])
        >>> ci.equals(ci5)
        False

        >>> ci6 = ci.set_categories(['b', 'c', 'a'])
        >>> ci.equals(ci6)
        True
        >>> ci_ordered = pd.CategoricalIndex(['a', 'b', 'c', 'a', 'b', 'c'],
        ...                                  ordered=True)
        >>> ci2_ordered = ci_ordered.set_categories(['b', 'c', 'a'])
        >>> ci_ordered.equals(ci2_ordered)
        False
        """
        if self.is_(other):
            return True

        if not isinstance(other, Index):
            return False

        try:
            other = self._is_dtype_compat(other)
        except (TypeError, ValueError):
            return False

        return self._data.equals(other)

    # --------------------------------------------------------------------
    # Rendering Methods

    @property
    def _formatter_func(self):
        return self.categories._formatter_func

    def _format_attrs(self):
        """
        Return a list of tuples of the (attr,formatted_value)
        """
        attrs: list[tuple[str, str | int | bool | None]]

        attrs = [
            (
                "categories",
                f"[{', '.join(self._data._repr_categories())}]",
            ),
            ("ordered", self.ordered),
        ]
        extra = super()._format_attrs()
        return attrs + extra

    # --------------------------------------------------------------------

    @property
    def inferred_type(self) -> str:
        return "categorical"

    @doc(Index.__contains__)
    def __contains__(self, key: Any) -> bool:
        # if key is a NaN, check if any NaN is in self.
        if is_valid_na_for_dtype(key, self.categories.dtype):
            return self.hasnans

        return contains(self, key, container=self._engine)

    def reindex(
        self, target, method=None, level=None, limit: int | None = None, tolerance=None
    ) -> tuple[Index, npt.NDArray[np.intp] | None]:
        """
        Create index with target's values (move/add/delete values as necessary)

        Returns
        -------
        new_index : pd.Index
            Resulting index
        indexer : np.ndarray[np.intp] or None
            Indices of output values in original index

        """
        if method is not None:
            raise NotImplementedError(
                "argument method is not implemented for CategoricalIndex.reindex"
            )
        if level is not None:
            raise NotImplementedError(
                "argument level is not implemented for CategoricalIndex.reindex"
            )
        if limit is not None:
            raise NotImplementedError(
                "argument limit is not implemented for CategoricalIndex.reindex"
            )
        return super().reindex(target)

    # --------------------------------------------------------------------
    # Indexing Methods

    def _maybe_cast_indexer(self, key) -> int:
        # GH#41933: we have to do this instead of self._data._validate_scalar
        #  because this will correctly get partial-indexing on Interval categories
        try:
            return self._data._unbox_scalar(key)
        except KeyError:
            if is_valid_na_for_dtype(key, self.categories.dtype):
                return -1
            raise

    def _maybe_cast_listlike_indexer(self, values) -> CategoricalIndex:
        if isinstance(values, CategoricalIndex):
            values = values._data
        if isinstance(values, Categorical):
            # Indexing on codes is more efficient if categories are the same,
            #  so we can apply some optimizations based on the degree of
            #  dtype-matching.
            cat = self._data._encode_with_my_categories(values)
            codes = cat._codes
        else:
            codes = self.categories.get_indexer(values)
            codes = codes.astype(self.codes.dtype, copy=False)
            cat = self._data._from_backing_data(codes)
        return type(self)._simple_new(cat)

    # --------------------------------------------------------------------

    def _is_comparable_dtype(self, dtype: DtypeObj) -> bool:
        return self.categories._is_comparable_dtype(dtype)

    def map(self, mapper, na_action: Literal["ignore"] | None = None):
        """
        Map values using input an input mapping or function.

        Maps the values (their categories, not the codes) of the index to new
        categories. If the mapping correspondence is one-to-one the result is a
        :class:`~pandas.CategoricalIndex` which has the same order property as
        the original, otherwise an :class:`~pandas.Index` is returned.

        If a `dict` or :class:`~pandas.Series` is used any unmapped category is
        mapped to `NaN`. Note that if this happens an :class:`~pandas.Index`
        will be returned.

        Parameters
        ----------
        mapper : function, dict, or Series
            Mapping correspondence.

        Returns
        -------
        pandas.CategoricalIndex or pandas.Index
            Mapped index.

        See Also
        --------
        Index.map : Apply a mapping correspondence on an
            :class:`~pandas.Index`.
        Series.map : Apply a mapping correspondence on a
            :class:`~pandas.Series`.
        Series.apply : Apply more complex functions on a
            :class:`~pandas.Series`.

        Examples
        --------
        >>> idx = pd.CategoricalIndex(['a', 'b', 'c'])
        >>> idx
        CategoricalIndex(['a', 'b', 'c'], categories=['a', 'b', 'c'],
                          ordered=False, dtype='category')
        >>> idx.map(lambda x: x.upper())
        CategoricalIndex(['A', 'B', 'C'], categories=['A', 'B', 'C'],
                         ordered=False, dtype='category')
        >>> idx.map({'a': 'first', 'b': 'second', 'c': 'third'})
        CategoricalIndex(['first', 'second', 'third'], categories=['first',
                         'second', 'third'], ordered=False, dtype='category')

        If the mapping is one-to-one the ordering of the categories is
        preserved:

        >>> idx = pd.CategoricalIndex(['a', 'b', 'c'], ordered=True)
        >>> idx
        CategoricalIndex(['a', 'b', 'c'], categories=['a', 'b', 'c'],
                         ordered=True, dtype='category')
        >>> idx.map({'a': 3, 'b': 2, 'c': 1})
        CategoricalIndex([3, 2, 1], categories=[3, 2, 1], ordered=True,
                         dtype='category')

        If the mapping is not one-to-one an :class:`~pandas.Index` is returned:

        >>> idx.map({'a': 'first', 'b': 'second', 'c': 'first'})
        Index(['first', 'second', 'first'], dtype='object')

        If a `dict` is used, all unmapped categories are mapped to `NaN` and
        the result is an :class:`~pandas.Index`:

        >>> idx.map({'a': 'first', 'b': 'second'})
        Index(['first', 'second', nan], dtype='object')
        """
        mapped = self._values.map(mapper, na_action=na_action)
        return Index(mapped, name=self.name)

    def _concat(self, to_concat: list[Index], name: Hashable) -> Index:
        # if calling index is category, don't check dtype of others
        try:
            cat = Categorical._concat_same_type(
                [self._is_dtype_compat(c) for c in to_concat]
            )
        except TypeError:
            # not all to_concat elements are among our categories (or NA)

            res = concat_compat([x._values for x in to_concat])
            return Index(res, name=name)
        else:
            return type(self)._simple_new(cat, name=name)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\vector\functions.py ===
from sympy.vector.coordsysrect import CoordSys3D
from sympy.vector.deloperator import Del
from sympy.vector.scalar import BaseScalar
from sympy.vector.vector import Vector, BaseVector
from sympy.vector.operators import gradient, curl, divergence
from sympy.core.function import diff
from sympy.core.singleton import S
from sympy.integrals.integrals import integrate
from sympy.core import sympify
from sympy.vector.dyadic import Dyadic


def express(expr, system, system2=None, variables=False):
    """
    Global function for 'express' functionality.

    Re-expresses a Vector, Dyadic or scalar(sympyfiable) in the given
    coordinate system.

    If 'variables' is True, then the coordinate variables (base scalars)
    of other coordinate systems present in the vector/scalar field or
    dyadic are also substituted in terms of the base scalars of the
    given system.

    Parameters
    ==========

    expr : Vector/Dyadic/scalar(sympyfiable)
        The expression to re-express in CoordSys3D 'system'

    system: CoordSys3D
        The coordinate system the expr is to be expressed in

    system2: CoordSys3D
        The other coordinate system required for re-expression
        (only for a Dyadic Expr)

    variables : boolean
        Specifies whether to substitute the coordinate variables present
        in expr, in terms of those of parameter system

    Examples
    ========

    >>> from sympy.vector import CoordSys3D
    >>> from sympy import Symbol, cos, sin
    >>> N = CoordSys3D('N')
    >>> q = Symbol('q')
    >>> B = N.orient_new_axis('B', q, N.k)
    >>> from sympy.vector import express
    >>> express(B.i, N)
    (cos(q))*N.i + (sin(q))*N.j
    >>> express(N.x, B, variables=True)
    B.x*cos(q) - B.y*sin(q)
    >>> d = N.i.outer(N.i)
    >>> express(d, B, N) == (cos(q))*(B.i|N.i) + (-sin(q))*(B.j|N.i)
    True

    """

    if expr in (0, Vector.zero):
        return expr

    if not isinstance(system, CoordSys3D):
        raise TypeError("system should be a CoordSys3D \
                        instance")

    if isinstance(expr, Vector):
        if system2 is not None:
            raise ValueError("system2 should not be provided for \
                                Vectors")
        # Given expr is a Vector
        if variables:
            # If variables attribute is True, substitute
            # the coordinate variables in the Vector
            system_list = {x.system for x in expr.atoms(BaseScalar, BaseVector)} - {system}
            subs_dict = {}
            for f in system_list:
                subs_dict.update(f.scalar_map(system))
            expr = expr.subs(subs_dict)
        # Re-express in this coordinate system
        outvec = Vector.zero
        parts = expr.separate()
        for x in parts:
            if x != system:
                temp = system.rotation_matrix(x) * parts[x].to_matrix(x)
                outvec += matrix_to_vector(temp, system)
            else:
                outvec += parts[x]
        return outvec

    elif isinstance(expr, Dyadic):
        if system2 is None:
            system2 = system
        if not isinstance(system2, CoordSys3D):
            raise TypeError("system2 should be a CoordSys3D \
                            instance")
        outdyad = Dyadic.zero
        var = variables
        for k, v in expr.components.items():
            outdyad += (express(v, system, variables=var) *
                        (express(k.args[0], system, variables=var) |
                         express(k.args[1], system2, variables=var)))

        return outdyad

    else:
        if system2 is not None:
            raise ValueError("system2 should not be provided for \
                                Vectors")
        if variables:
            # Given expr is a scalar field
            system_set = set()
            expr = sympify(expr)
            # Substitute all the coordinate variables
            for x in expr.atoms(BaseScalar):
                if x.system != system:
                    system_set.add(x.system)
            subs_dict = {}
            for f in system_set:
                subs_dict.update(f.scalar_map(system))
            return expr.subs(subs_dict)
        return expr


def directional_derivative(field, direction_vector):
    """
    Returns the directional derivative of a scalar or vector field computed
    along a given vector in coordinate system which parameters are expressed.

    Parameters
    ==========

    field : Vector or Scalar
        The scalar or vector field to compute the directional derivative of

    direction_vector : Vector
        The vector to calculated directional derivative along them.


    Examples
    ========

    >>> from sympy.vector import CoordSys3D, directional_derivative
    >>> R = CoordSys3D('R')
    >>> f1 = R.x*R.y*R.z
    >>> v1 = 3*R.i + 4*R.j + R.k
    >>> directional_derivative(f1, v1)
    R.x*R.y + 4*R.x*R.z + 3*R.y*R.z
    >>> f2 = 5*R.x**2*R.z
    >>> directional_derivative(f2, v1)
    5*R.x**2 + 30*R.x*R.z

    """
    from sympy.vector.operators import _get_coord_systems
    coord_sys = _get_coord_systems(field)
    if len(coord_sys) > 0:
        # TODO: This gets a random coordinate system in case of multiple ones:
        coord_sys = next(iter(coord_sys))
        field = express(field, coord_sys, variables=True)
        i, j, k = coord_sys.base_vectors()
        x, y, z = coord_sys.base_scalars()
        out = Vector.dot(direction_vector, i) * diff(field, x)
        out += Vector.dot(direction_vector, j) * diff(field, y)
        out += Vector.dot(direction_vector, k) * diff(field, z)
        if out == 0 and isinstance(field, Vector):
            out = Vector.zero
        return out
    elif isinstance(field, Vector):
        return Vector.zero
    else:
        return S.Zero


def laplacian(expr):
    """
    Return the laplacian of the given field computed in terms of
    the base scalars of the given coordinate system.

    Parameters
    ==========

    expr : SymPy Expr or Vector
        expr denotes a scalar or vector field.

    Examples
    ========

    >>> from sympy.vector import CoordSys3D, laplacian
    >>> R = CoordSys3D('R')
    >>> f = R.x**2*R.y**5*R.z
    >>> laplacian(f)
    20*R.x**2*R.y**3*R.z + 2*R.y**5*R.z
    >>> f = R.x**2*R.i + R.y**3*R.j + R.z**4*R.k
    >>> laplacian(f)
    2*R.i + 6*R.y*R.j + 12*R.z**2*R.k

    """

    delop = Del()
    if expr.is_Vector:
        return (gradient(divergence(expr)) - curl(curl(expr))).doit()
    return delop.dot(delop(expr)).doit()


def is_conservative(field):
    """
    Checks if a field is conservative.

    Parameters
    ==========

    field : Vector
        The field to check for conservative property

    Examples
    ========

    >>> from sympy.vector import CoordSys3D
    >>> from sympy.vector import is_conservative
    >>> R = CoordSys3D('R')
    >>> is_conservative(R.y*R.z*R.i + R.x*R.z*R.j + R.x*R.y*R.k)
    True
    >>> is_conservative(R.z*R.j)
    False

    """

    # Field is conservative irrespective of system
    # Take the first coordinate system in the result of the
    # separate method of Vector
    if not isinstance(field, Vector):
        raise TypeError("field should be a Vector")
    if field == Vector.zero:
        return True
    return curl(field).simplify() == Vector.zero


def is_solenoidal(field):
    """
    Checks if a field is solenoidal.

    Parameters
    ==========

    field : Vector
        The field to check for solenoidal property

    Examples
    ========

    >>> from sympy.vector import CoordSys3D
    >>> from sympy.vector import is_solenoidal
    >>> R = CoordSys3D('R')
    >>> is_solenoidal(R.y*R.z*R.i + R.x*R.z*R.j + R.x*R.y*R.k)
    True
    >>> is_solenoidal(R.y * R.j)
    False

    """

    # Field is solenoidal irrespective of system
    # Take the first coordinate system in the result of the
    # separate method in Vector
    if not isinstance(field, Vector):
        raise TypeError("field should be a Vector")
    if field == Vector.zero:
        return True
    return divergence(field).simplify() is S.Zero


def scalar_potential(field, coord_sys):
    """
    Returns the scalar potential function of a field in a given
    coordinate system (without the added integration constant).

    Parameters
    ==========

    field : Vector
        The vector field whose scalar potential function is to be
        calculated

    coord_sys : CoordSys3D
        The coordinate system to do the calculation in

    Examples
    ========

    >>> from sympy.vector import CoordSys3D
    >>> from sympy.vector import scalar_potential, gradient
    >>> R = CoordSys3D('R')
    >>> scalar_potential(R.k, R) == R.z
    True
    >>> scalar_field = 2*R.x**2*R.y*R.z
    >>> grad_field = gradient(scalar_field)
    >>> scalar_potential(grad_field, R)
    2*R.x**2*R.y*R.z

    """

    # Check whether field is conservative
    if not is_conservative(field):
        raise ValueError("Field is not conservative")
    if field == Vector.zero:
        return S.Zero
    # Express the field exntirely in coord_sys
    # Substitute coordinate variables also
    if not isinstance(coord_sys, CoordSys3D):
        raise TypeError("coord_sys must be a CoordSys3D")
    field = express(field, coord_sys, variables=True)
    dimensions = coord_sys.base_vectors()
    scalars = coord_sys.base_scalars()
    # Calculate scalar potential function
    temp_function = integrate(field.dot(dimensions[0]), scalars[0])
    for i, dim in enumerate(dimensions[1:]):
        partial_diff = diff(temp_function, scalars[i + 1])
        partial_diff = field.dot(dim) - partial_diff
        temp_function += integrate(partial_diff, scalars[i + 1])
    return temp_function


def scalar_potential_difference(field, coord_sys, point1, point2):
    """
    Returns the scalar potential difference between two points in a
    certain coordinate system, wrt a given field.

    If a scalar field is provided, its values at the two points are
    considered. If a conservative vector field is provided, the values
    of its scalar potential function at the two points are used.

    Returns (potential at point2) - (potential at point1)

    The position vectors of the two Points are calculated wrt the
    origin of the coordinate system provided.

    Parameters
    ==========

    field : Vector/Expr
        The field to calculate wrt

    coord_sys : CoordSys3D
        The coordinate system to do the calculations in

    point1 : Point
        The initial Point in given coordinate system

    position2 : Point
        The second Point in the given coordinate system

    Examples
    ========

    >>> from sympy.vector import CoordSys3D
    >>> from sympy.vector import scalar_potential_difference
    >>> R = CoordSys3D('R')
    >>> P = R.origin.locate_new('P', R.x*R.i + R.y*R.j + R.z*R.k)
    >>> vectfield = 4*R.x*R.y*R.i + 2*R.x**2*R.j
    >>> scalar_potential_difference(vectfield, R, R.origin, P)
    2*R.x**2*R.y
    >>> Q = R.origin.locate_new('O', 3*R.i + R.j + 2*R.k)
    >>> scalar_potential_difference(vectfield, R, P, Q)
    -2*R.x**2*R.y + 18

    """

    if not isinstance(coord_sys, CoordSys3D):
        raise TypeError("coord_sys must be a CoordSys3D")
    if isinstance(field, Vector):
        # Get the scalar potential function
        scalar_fn = scalar_potential(field, coord_sys)
    else:
        # Field is a scalar
        scalar_fn = field
    # Express positions in required coordinate system
    origin = coord_sys.origin
    position1 = express(point1.position_wrt(origin), coord_sys,
                        variables=True)
    position2 = express(point2.position_wrt(origin), coord_sys,
                        variables=True)
    # Get the two positions as substitution dicts for coordinate variables
    subs_dict1 = {}
    subs_dict2 = {}
    scalars = coord_sys.base_scalars()
    for i, x in enumerate(coord_sys.base_vectors()):
        subs_dict1[scalars[i]] = x.dot(position1)
        subs_dict2[scalars[i]] = x.dot(position2)
    return scalar_fn.subs(subs_dict2) - scalar_fn.subs(subs_dict1)


def matrix_to_vector(matrix, system):
    """
    Converts a vector in matrix form to a Vector instance.

    It is assumed that the elements of the Matrix represent the
    measure numbers of the components of the vector along basis
    vectors of 'system'.

    Parameters
    ==========

    matrix : SymPy Matrix, Dimensions: (3, 1)
        The matrix to be converted to a vector

    system : CoordSys3D
        The coordinate system the vector is to be defined in

    Examples
    ========

    >>> from sympy import ImmutableMatrix as Matrix
    >>> m = Matrix([1, 2, 3])
    >>> from sympy.vector import CoordSys3D, matrix_to_vector
    >>> C = CoordSys3D('C')
    >>> v = matrix_to_vector(m, C)
    >>> v
    C.i + 2*C.j + 3*C.k
    >>> v.to_matrix(C) == m
    True

    """

    outvec = Vector.zero
    vects = system.base_vectors()
    for i, x in enumerate(matrix):
        outvec += x * vects[i]
    return outvec


def _path(from_object, to_object):
    """
    Calculates the 'path' of objects starting from 'from_object'
    to 'to_object', along with the index of the first common
    ancestor in the tree.

    Returns (index, list) tuple.
    """

    if from_object._root != to_object._root:
        raise ValueError("No connecting path found between " +
                         str(from_object) + " and " + str(to_object))

    other_path = []
    obj = to_object
    while obj._parent is not None:
        other_path.append(obj)
        obj = obj._parent
    other_path.append(obj)
    object_set = set(other_path)
    from_path = []
    obj = from_object
    while obj not in object_set:
        from_path.append(obj)
        obj = obj._parent
    index = len(from_path)
    from_path.extend(other_path[other_path.index(obj)::-1])
    return index, from_path


def orthogonalize(*vlist, orthonormal=False):
    """
    Takes a sequence of independent vectors and orthogonalizes them
    using the Gram - Schmidt process. Returns a list of
    orthogonal or orthonormal vectors.

    Parameters
    ==========

    vlist : sequence of independent vectors to be made orthogonal.

    orthonormal : Optional parameter
                  Set to True if the vectors returned should be
                  orthonormal.
                  Default: False

    Examples
    ========

    >>> from sympy.vector.coordsysrect import CoordSys3D
    >>> from sympy.vector.functions import orthogonalize
    >>> C = CoordSys3D('C')
    >>> i, j, k = C.base_vectors()
    >>> v1 = i + 2*j
    >>> v2 = 2*i + 3*j
    >>> orthogonalize(v1, v2)
    [C.i + 2*C.j, 2/5*C.i + (-1/5)*C.j]

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Gram-Schmidt_process

    """

    if not all(isinstance(vec, Vector) for vec in vlist):
        raise TypeError('Each element must be of Type Vector')

    ortho_vlist = []
    for i, term in enumerate(vlist):
        for j in range(i):
            term -= ortho_vlist[j].projection(vlist[i])
        # TODO : The following line introduces a performance issue
        # and needs to be changed once a good solution for issue #10279 is
        # found.
        if term.equals(Vector.zero):
            raise ValueError("Vector set not linearly independent")
        ortho_vlist.append(term)

    if orthonormal:
        ortho_vlist = [vec.normalize() for vec in ortho_vlist]

    return ortho_vlist

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_file_io.py ===
from __future__ import annotations

import io
from collections.abc import Callable, Iterable
from functools import partial
from typing import (
    IO,
    TYPE_CHECKING,
    Any,
    AnyStr,
    BinaryIO,
    Generic,
    TypeVar,
    Union,
    overload,
)

import trio

from ._util import async_wraps
from .abc import AsyncResource

if TYPE_CHECKING:
    from _typeshed import (
        OpenBinaryMode,
        OpenBinaryModeReading,
        OpenBinaryModeUpdating,
        OpenBinaryModeWriting,
        OpenTextMode,
        StrOrBytesPath,
    )
    from typing_extensions import Literal

    from ._sync import CapacityLimiter

# This list is also in the docs, make sure to keep them in sync
_FILE_SYNC_ATTRS: set[str] = {
    "closed",
    "encoding",
    "errors",
    "fileno",
    "isatty",
    "newlines",
    "readable",
    "seekable",
    "writable",
    # not defined in *IOBase:
    "buffer",
    "raw",
    "line_buffering",
    "closefd",
    "name",
    "mode",
    "getvalue",
    "getbuffer",
}

# This list is also in the docs, make sure to keep them in sync
_FILE_ASYNC_METHODS: set[str] = {
    "flush",
    "read",
    "read1",
    "readall",
    "readinto",
    "readline",
    "readlines",
    "seek",
    "tell",
    "truncate",
    "write",
    "writelines",
    # not defined in *IOBase:
    "readinto1",
    "peek",
}


FileT = TypeVar("FileT")
FileT_co = TypeVar("FileT_co", covariant=True)
T = TypeVar("T")
T_co = TypeVar("T_co", covariant=True)
T_contra = TypeVar("T_contra", contravariant=True)
AnyStr_co = TypeVar("AnyStr_co", str, bytes, covariant=True)
AnyStr_contra = TypeVar("AnyStr_contra", str, bytes, contravariant=True)

# This is a little complicated. IO objects have a lot of methods, and which are available on
# different types varies wildly. We want to match the interface of whatever file we're wrapping.
# This pile of protocols each has one sync method/property, meaning they're going to be compatible
# with a file class that supports that method/property. The ones parameterized with AnyStr take
# either str or bytes depending.

# The wrapper is then a generic class, where the typevar is set to the type of the sync file we're
# wrapping. For generics, adding a type to self has a special meaning - properties/methods can be
# conditional - it's only valid to call them if the object you're accessing them on is compatible
# with that type hint. By using the protocols, the type checker will be checking to see if the
# wrapped type has that method, and only allow the methods that do to be called. We can then alter
# the signature however it needs to match runtime behaviour.
# More info: https://mypy.readthedocs.io/en/stable/more_types.html#advanced-uses-of-self-types
if TYPE_CHECKING:
    from typing_extensions import Buffer, Protocol

    # fmt: off

    class _HasClosed(Protocol):
        @property
        def closed(self) -> bool: ...

    class _HasEncoding(Protocol):
        @property
        def encoding(self) -> str: ...

    class _HasErrors(Protocol):
        @property
        def errors(self) -> str | None: ...

    class _HasFileNo(Protocol):
        def fileno(self) -> int: ...

    class _HasIsATTY(Protocol):
        def isatty(self) -> bool: ...

    class _HasNewlines(Protocol[T_co]):
        # Type varies here - documented to be None, tuple of strings, strings. Typeshed uses Any.
        @property
        def newlines(self) -> T_co: ...

    class _HasReadable(Protocol):
        def readable(self) -> bool: ...

    class _HasSeekable(Protocol):
        def seekable(self) -> bool: ...

    class _HasWritable(Protocol):
        def writable(self) -> bool: ...

    class _HasBuffer(Protocol):
        @property
        def buffer(self) -> BinaryIO: ...

    class _HasRaw(Protocol):
        @property
        def raw(self) -> io.RawIOBase: ...

    class _HasLineBuffering(Protocol):
        @property
        def line_buffering(self) -> bool: ...

    class _HasCloseFD(Protocol):
        @property
        def closefd(self) -> bool: ...

    class _HasName(Protocol):
        @property
        def name(self) -> str: ...

    class _HasMode(Protocol):
        @property
        def mode(self) -> str: ...

    class _CanGetValue(Protocol[AnyStr_co]):
        def getvalue(self) -> AnyStr_co: ...

    class _CanGetBuffer(Protocol):
        def getbuffer(self) -> memoryview: ...

    class _CanFlush(Protocol):
        def flush(self) -> None: ...

    class _CanRead(Protocol[AnyStr_co]):
        def read(self, size: int | None = ..., /) -> AnyStr_co: ...

    class _CanRead1(Protocol):
        def read1(self, size: int | None = ..., /) -> bytes: ...

    class _CanReadAll(Protocol[AnyStr_co]):
        def readall(self) -> AnyStr_co: ...

    class _CanReadInto(Protocol):
        def readinto(self, buf: Buffer, /) -> int | None: ...

    class _CanReadInto1(Protocol):
        def readinto1(self, buffer: Buffer, /) -> int: ...

    class _CanReadLine(Protocol[AnyStr_co]):
        def readline(self, size: int = ..., /) -> AnyStr_co: ...

    class _CanReadLines(Protocol[AnyStr]):
        def readlines(self, hint: int = ..., /) -> list[AnyStr]: ...

    class _CanSeek(Protocol):
        def seek(self, target: int, whence: int = 0, /) -> int: ...

    class _CanTell(Protocol):
        def tell(self) -> int: ...

    class _CanTruncate(Protocol):
        def truncate(self, size: int | None = ..., /) -> int: ...

    class _CanWrite(Protocol[T_contra]):
        def write(self, data: T_contra, /) -> int: ...

    class _CanWriteLines(Protocol[T_contra]):
        # The lines parameter varies for bytes/str, so use a typevar to make the async match.
        def writelines(self, lines: Iterable[T_contra], /) -> None: ...

    class _CanPeek(Protocol[AnyStr_co]):
        def peek(self, size: int = 0, /) -> AnyStr_co: ...

    class _CanDetach(Protocol[T_co]):
        # The T typevar will be the unbuffered/binary file this file wraps.
        def detach(self) -> T_co: ...

    class _CanClose(Protocol):
        def close(self) -> None: ...


# FileT needs to be covariant for the protocol trick to work - the real IO types are effectively a
# subtype of the protocols.
class AsyncIOWrapper(AsyncResource, Generic[FileT_co]):
    """A generic :class:`~io.IOBase` wrapper that implements the :term:`asynchronous
    file object` interface. Wrapped methods that could block are executed in
    :meth:`trio.to_thread.run_sync`.

    All properties and methods defined in :mod:`~io` are exposed by this
    wrapper, if they exist in the wrapped file object.
    """

    def __init__(self, file: FileT_co) -> None:
        self._wrapped = file

    @property
    def wrapped(self) -> FileT_co:
        """object: A reference to the wrapped file object"""

        return self._wrapped

    if not TYPE_CHECKING:

        def __getattr__(self, name: str) -> object:
            if name in _FILE_SYNC_ATTRS:
                return getattr(self._wrapped, name)
            if name in _FILE_ASYNC_METHODS:
                meth = getattr(self._wrapped, name)

                @async_wraps(self.__class__, self._wrapped.__class__, name)
                async def wrapper(
                    *args: Callable[..., T],
                    **kwargs: object | str | bool | CapacityLimiter | None,
                ) -> T:
                    func = partial(meth, *args, **kwargs)
                    return await trio.to_thread.run_sync(func)

                # cache the generated method
                setattr(self, name, wrapper)
                return wrapper

            raise AttributeError(name)

    def __dir__(self) -> Iterable[str]:
        attrs = set(super().__dir__())
        attrs.update(a for a in _FILE_SYNC_ATTRS if hasattr(self.wrapped, a))
        attrs.update(a for a in _FILE_ASYNC_METHODS if hasattr(self.wrapped, a))
        return attrs

    def __aiter__(self) -> AsyncIOWrapper[FileT_co]:
        return self

    async def __anext__(self: AsyncIOWrapper[_CanReadLine[AnyStr]]) -> AnyStr:
        line = await self.readline()
        if line:
            return line
        else:
            raise StopAsyncIteration

    async def detach(self: AsyncIOWrapper[_CanDetach[T]]) -> AsyncIOWrapper[T]:
        """Like :meth:`io.BufferedIOBase.detach`, but async.

        This also re-wraps the result in a new :term:`asynchronous file object`
        wrapper.

        """

        raw = await trio.to_thread.run_sync(self._wrapped.detach)
        return wrap_file(raw)

    async def aclose(self: AsyncIOWrapper[_CanClose]) -> None:
        """Like :meth:`io.IOBase.close`, but async.

        This is also shielded from cancellation; if a cancellation scope is
        cancelled, the wrapped file object will still be safely closed.

        """

        # ensure the underling file is closed during cancellation
        with trio.CancelScope(shield=True):
            await trio.to_thread.run_sync(self._wrapped.close)

        await trio.lowlevel.checkpoint_if_cancelled()

    if TYPE_CHECKING:
        # fmt: off
        # Based on typing.IO and io stubs.
        @property
        def closed(self: AsyncIOWrapper[_HasClosed]) -> bool: ...
        @property
        def encoding(self: AsyncIOWrapper[_HasEncoding]) -> str: ...
        @property
        def errors(self: AsyncIOWrapper[_HasErrors]) -> str | None: ...
        @property
        def newlines(self: AsyncIOWrapper[_HasNewlines[T]]) -> T: ...
        @property
        def buffer(self: AsyncIOWrapper[_HasBuffer]) -> BinaryIO: ...
        @property
        def raw(self: AsyncIOWrapper[_HasRaw]) -> io.RawIOBase: ...
        @property
        def line_buffering(self: AsyncIOWrapper[_HasLineBuffering]) -> int: ...
        @property
        def closefd(self: AsyncIOWrapper[_HasCloseFD]) -> bool: ...
        @property
        def name(self: AsyncIOWrapper[_HasName]) -> str: ...
        @property
        def mode(self: AsyncIOWrapper[_HasMode]) -> str: ...

        def fileno(self: AsyncIOWrapper[_HasFileNo]) -> int: ...
        def isatty(self: AsyncIOWrapper[_HasIsATTY]) -> bool: ...
        def readable(self: AsyncIOWrapper[_HasReadable]) -> bool: ...
        def seekable(self: AsyncIOWrapper[_HasSeekable]) -> bool: ...
        def writable(self: AsyncIOWrapper[_HasWritable]) -> bool: ...
        def getvalue(self: AsyncIOWrapper[_CanGetValue[AnyStr]]) -> AnyStr: ...
        def getbuffer(self: AsyncIOWrapper[_CanGetBuffer]) -> memoryview: ...
        async def flush(self: AsyncIOWrapper[_CanFlush]) -> None: ...
        async def read(self: AsyncIOWrapper[_CanRead[AnyStr]], size: int | None = -1, /) -> AnyStr: ...
        async def read1(self: AsyncIOWrapper[_CanRead1], size: int | None = -1, /) -> bytes: ...
        async def readall(self: AsyncIOWrapper[_CanReadAll[AnyStr]]) -> AnyStr: ...
        async def readinto(self: AsyncIOWrapper[_CanReadInto], buf: Buffer, /) -> int | None: ...
        async def readline(self: AsyncIOWrapper[_CanReadLine[AnyStr]], size: int = -1, /) -> AnyStr: ...
        async def readlines(self: AsyncIOWrapper[_CanReadLines[AnyStr]]) -> list[AnyStr]: ...
        async def seek(self: AsyncIOWrapper[_CanSeek], target: int, whence: int = 0, /) -> int: ...
        async def tell(self: AsyncIOWrapper[_CanTell]) -> int: ...
        async def truncate(self: AsyncIOWrapper[_CanTruncate], size: int | None = None, /) -> int: ...
        async def write(self: AsyncIOWrapper[_CanWrite[T]], data: T, /) -> int: ...
        async def writelines(self: AsyncIOWrapper[_CanWriteLines[T]], lines: Iterable[T], /) -> None: ...
        async def readinto1(self: AsyncIOWrapper[_CanReadInto1], buffer: Buffer, /) -> int: ...
        async def peek(self: AsyncIOWrapper[_CanPeek[AnyStr]], size: int = 0, /) -> AnyStr: ...


# Type hints are copied from builtin open.
_OpenFile = Union["StrOrBytesPath", int]
_Opener = Callable[[str, int], int]


@overload
async def open_file(
    file: _OpenFile,
    mode: OpenTextMode = "r",
    buffering: int = -1,
    encoding: str | None = None,
    errors: str | None = None,
    newline: str | None = None,
    closefd: bool = True,
    opener: _Opener | None = None,
) -> AsyncIOWrapper[io.TextIOWrapper]: ...


@overload
async def open_file(
    file: _OpenFile,
    mode: OpenBinaryMode,
    buffering: Literal[0],
    encoding: None = None,
    errors: None = None,
    newline: None = None,
    closefd: bool = True,
    opener: _Opener | None = None,
) -> AsyncIOWrapper[io.FileIO]: ...


@overload
async def open_file(
    file: _OpenFile,
    mode: OpenBinaryModeUpdating,
    buffering: Literal[-1, 1] = -1,
    encoding: None = None,
    errors: None = None,
    newline: None = None,
    closefd: bool = True,
    opener: _Opener | None = None,
) -> AsyncIOWrapper[io.BufferedRandom]: ...


@overload
async def open_file(
    file: _OpenFile,
    mode: OpenBinaryModeWriting,
    buffering: Literal[-1, 1] = -1,
    encoding: None = None,
    errors: None = None,
    newline: None = None,
    closefd: bool = True,
    opener: _Opener | None = None,
) -> AsyncIOWrapper[io.BufferedWriter]: ...


@overload
async def open_file(
    file: _OpenFile,
    mode: OpenBinaryModeReading,
    buffering: Literal[-1, 1] = -1,
    encoding: None = None,
    errors: None = None,
    newline: None = None,
    closefd: bool = True,
    opener: _Opener | None = None,
) -> AsyncIOWrapper[io.BufferedReader]: ...


@overload
async def open_file(
    file: _OpenFile,
    mode: OpenBinaryMode,
    buffering: int,
    encoding: None = None,
    errors: None = None,
    newline: None = None,
    closefd: bool = True,
    opener: _Opener | None = None,
) -> AsyncIOWrapper[BinaryIO]: ...


@overload
async def open_file(  # type: ignore[explicit-any, misc]  # Any usage matches builtins.open().
    file: _OpenFile,
    mode: str,
    buffering: int = -1,
    encoding: str | None = None,
    errors: str | None = None,
    newline: str | None = None,
    closefd: bool = True,
    opener: _Opener | None = None,
) -> AsyncIOWrapper[IO[Any]]: ...


async def open_file(
    file: _OpenFile,
    mode: str = "r",
    buffering: int = -1,
    encoding: str | None = None,
    errors: str | None = None,
    newline: str | None = None,
    closefd: bool = True,
    opener: _Opener | None = None,
) -> AsyncIOWrapper[object]:
    """Asynchronous version of :func:`open`.

    Returns:
        An :term:`asynchronous file object`

    Example::

        async with await trio.open_file(filename) as f:
            async for line in f:
                pass

        assert f.closed

    See also:
      :func:`trio.Path.open`

    """
    file_ = wrap_file(
        await trio.to_thread.run_sync(
            io.open,
            file,
            mode,
            buffering,
            encoding,
            errors,
            newline,
            closefd,
            opener,
        ),
    )
    return file_


def wrap_file(file: FileT) -> AsyncIOWrapper[FileT]:
    """This wraps any file object in a wrapper that provides an asynchronous
    file object interface.

    Args:
        file: a :term:`file object`

    Returns:
        An :term:`asynchronous file object` that wraps ``file``

    Example::

        async_file = trio.wrap_file(StringIO('asdf'))

        assert await async_file.read() == 'asdf'

    """

    def has(attr: str) -> bool:
        return hasattr(file, attr) and callable(getattr(file, attr))

    if not (has("close") and (has("read") or has("write"))):
        raise TypeError(
            f"{file} does not implement required duck-file methods: "
            "close and (read or write)",
        )

    return AsyncIOWrapper(file)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\blinker\base.py ===
from __future__ import annotations

import collections.abc as c
import sys
import typing as t
import weakref
from collections import defaultdict
from contextlib import contextmanager
from functools import cached_property
from inspect import iscoroutinefunction

from ._utilities import make_id
from ._utilities import make_ref
from ._utilities import Symbol

F = t.TypeVar("F", bound=c.Callable[..., t.Any])

ANY = Symbol("ANY")
"""Symbol for "any sender"."""

ANY_ID = 0


class Signal:
    """A notification emitter.

    :param doc: The docstring for the signal.
    """

    ANY = ANY
    """An alias for the :data:`~blinker.ANY` sender symbol."""

    set_class: type[set[t.Any]] = set
    """The set class to use for tracking connected receivers and senders.
    Python's ``set`` is unordered. If receivers must be dispatched in the order
    they were connected, an ordered set implementation can be used.

    .. versionadded:: 1.7
    """

    @cached_property
    def receiver_connected(self) -> Signal:
        """Emitted at the end of each :meth:`connect` call.

        The signal sender is the signal instance, and the :meth:`connect`
        arguments are passed through: ``receiver``, ``sender``, and ``weak``.

        .. versionadded:: 1.2
        """
        return Signal(doc="Emitted after a receiver connects.")

    @cached_property
    def receiver_disconnected(self) -> Signal:
        """Emitted at the end of each :meth:`disconnect` call.

        The sender is the signal instance, and the :meth:`disconnect` arguments
        are passed through: ``receiver`` and ``sender``.

        This signal is emitted **only** when :meth:`disconnect` is called
        explicitly. This signal cannot be emitted by an automatic disconnect
        when a weakly referenced receiver or sender goes out of scope, as the
        instance is no longer be available to be used as the sender for this
        signal.

        An alternative approach is available by subscribing to
        :attr:`receiver_connected` and setting up a custom weakref cleanup
        callback on weak receivers and senders.

        .. versionadded:: 1.2
        """
        return Signal(doc="Emitted after a receiver disconnects.")

    def __init__(self, doc: str | None = None) -> None:
        if doc:
            self.__doc__ = doc

        self.receivers: dict[
            t.Any, weakref.ref[c.Callable[..., t.Any]] | c.Callable[..., t.Any]
        ] = {}
        """The map of connected receivers. Useful to quickly check if any
        receivers are connected to the signal: ``if s.receivers:``. The
        structure and data is not part of the public API, but checking its
        boolean value is.
        """

        self.is_muted: bool = False
        self._by_receiver: dict[t.Any, set[t.Any]] = defaultdict(self.set_class)
        self._by_sender: dict[t.Any, set[t.Any]] = defaultdict(self.set_class)
        self._weak_senders: dict[t.Any, weakref.ref[t.Any]] = {}

    def connect(self, receiver: F, sender: t.Any = ANY, weak: bool = True) -> F:
        """Connect ``receiver`` to be called when the signal is sent by
        ``sender``.

        :param receiver: The callable to call when :meth:`send` is called with
            the given ``sender``, passing ``sender`` as a positional argument
            along with any extra keyword arguments.
        :param sender: Any object or :data:`ANY`. ``receiver`` will only be
            called when :meth:`send` is called with this sender. If ``ANY``, the
            receiver will be called for any sender. A receiver may be connected
            to multiple senders by calling :meth:`connect` multiple times.
        :param weak: Track the receiver with a :mod:`weakref`. The receiver will
            be automatically disconnected when it is garbage collected. When
            connecting a receiver defined within a function, set to ``False``,
            otherwise it will be disconnected when the function scope ends.
        """
        receiver_id = make_id(receiver)
        sender_id = ANY_ID if sender is ANY else make_id(sender)

        if weak:
            self.receivers[receiver_id] = make_ref(
                receiver, self._make_cleanup_receiver(receiver_id)
            )
        else:
            self.receivers[receiver_id] = receiver

        self._by_sender[sender_id].add(receiver_id)
        self._by_receiver[receiver_id].add(sender_id)

        if sender is not ANY and sender_id not in self._weak_senders:
            # store a cleanup for weakref-able senders
            try:
                self._weak_senders[sender_id] = make_ref(
                    sender, self._make_cleanup_sender(sender_id)
                )
            except TypeError:
                pass

        if "receiver_connected" in self.__dict__ and self.receiver_connected.receivers:
            try:
                self.receiver_connected.send(
                    self, receiver=receiver, sender=sender, weak=weak
                )
            except TypeError:
                # TODO no explanation or test for this
                self.disconnect(receiver, sender)
                raise

        return receiver

    def connect_via(self, sender: t.Any, weak: bool = False) -> c.Callable[[F], F]:
        """Connect the decorated function to be called when the signal is sent
        by ``sender``.

        The decorated function will be called when :meth:`send` is called with
        the given ``sender``, passing ``sender`` as a positional argument along
        with any extra keyword arguments.

        :param sender: Any object or :data:`ANY`. ``receiver`` will only be
            called when :meth:`send` is called with this sender. If ``ANY``, the
            receiver will be called for any sender. A receiver may be connected
            to multiple senders by calling :meth:`connect` multiple times.
        :param weak: Track the receiver with a :mod:`weakref`. The receiver will
            be automatically disconnected when it is garbage collected. When
            connecting a receiver defined within a function, set to ``False``,
            otherwise it will be disconnected when the function scope ends.=

        .. versionadded:: 1.1
        """

        def decorator(fn: F) -> F:
            self.connect(fn, sender, weak)
            return fn

        return decorator

    @contextmanager
    def connected_to(
        self, receiver: c.Callable[..., t.Any], sender: t.Any = ANY
    ) -> c.Generator[None, None, None]:
        """A context manager that temporarily connects ``receiver`` to the
        signal while a ``with`` block executes. When the block exits, the
        receiver is disconnected. Useful for tests.

        :param receiver: The callable to call when :meth:`send` is called with
            the given ``sender``, passing ``sender`` as a positional argument
            along with any extra keyword arguments.
        :param sender: Any object or :data:`ANY`. ``receiver`` will only be
            called when :meth:`send` is called with this sender. If ``ANY``, the
            receiver will be called for any sender.

        .. versionadded:: 1.1
        """
        self.connect(receiver, sender=sender, weak=False)

        try:
            yield None
        finally:
            self.disconnect(receiver)

    @contextmanager
    def muted(self) -> c.Generator[None, None, None]:
        """A context manager that temporarily disables the signal. No receivers
        will be called if the signal is sent, until the ``with`` block exits.
        Useful for tests.
        """
        self.is_muted = True

        try:
            yield None
        finally:
            self.is_muted = False

    def send(
        self,
        sender: t.Any | None = None,
        /,
        *,
        _async_wrapper: c.Callable[
            [c.Callable[..., c.Coroutine[t.Any, t.Any, t.Any]]], c.Callable[..., t.Any]
        ]
        | None = None,
        **kwargs: t.Any,
    ) -> list[tuple[c.Callable[..., t.Any], t.Any]]:
        """Call all receivers that are connected to the given ``sender``
        or :data:`ANY`. Each receiver is called with ``sender`` as a positional
        argument along with any extra keyword arguments. Return a list of
        ``(receiver, return value)`` tuples.

        The order receivers are called is undefined, but can be influenced by
        setting :attr:`set_class`.

        If a receiver raises an exception, that exception will propagate up.
        This makes debugging straightforward, with an assumption that correctly
        implemented receivers will not raise.

        :param sender: Call receivers connected to this sender, in addition to
            those connected to :data:`ANY`.
        :param _async_wrapper: Will be called on any receivers that are async
            coroutines to turn them into sync callables. For example, could run
            the receiver with an event loop.
        :param kwargs: Extra keyword arguments to pass to each receiver.

        .. versionchanged:: 1.7
            Added the ``_async_wrapper`` argument.
        """
        if self.is_muted:
            return []

        results = []

        for receiver in self.receivers_for(sender):
            if iscoroutinefunction(receiver):
                if _async_wrapper is None:
                    raise RuntimeError("Cannot send to a coroutine function.")

                result = _async_wrapper(receiver)(sender, **kwargs)
            else:
                result = receiver(sender, **kwargs)

            results.append((receiver, result))

        return results

    async def send_async(
        self,
        sender: t.Any | None = None,
        /,
        *,
        _sync_wrapper: c.Callable[
            [c.Callable[..., t.Any]], c.Callable[..., c.Coroutine[t.Any, t.Any, t.Any]]
        ]
        | None = None,
        **kwargs: t.Any,
    ) -> list[tuple[c.Callable[..., t.Any], t.Any]]:
        """Await all receivers that are connected to the given ``sender``
        or :data:`ANY`. Each receiver is called with ``sender`` as a positional
        argument along with any extra keyword arguments. Return a list of
        ``(receiver, return value)`` tuples.

        The order receivers are called is undefined, but can be influenced by
        setting :attr:`set_class`.

        If a receiver raises an exception, that exception will propagate up.
        This makes debugging straightforward, with an assumption that correctly
        implemented receivers will not raise.

        :param sender: Call receivers connected to this sender, in addition to
            those connected to :data:`ANY`.
        :param _sync_wrapper: Will be called on any receivers that are sync
            callables to turn them into async coroutines. For example,
            could call the receiver in a thread.
        :param kwargs: Extra keyword arguments to pass to each receiver.

        .. versionadded:: 1.7
        """
        if self.is_muted:
            return []

        results = []

        for receiver in self.receivers_for(sender):
            if not iscoroutinefunction(receiver):
                if _sync_wrapper is None:
                    raise RuntimeError("Cannot send to a non-coroutine function.")

                result = await _sync_wrapper(receiver)(sender, **kwargs)
            else:
                result = await receiver(sender, **kwargs)

            results.append((receiver, result))

        return results

    def has_receivers_for(self, sender: t.Any) -> bool:
        """Check if there is at least one receiver that will be called with the
        given ``sender``. A receiver connected to :data:`ANY` will always be
        called, regardless of sender. Does not check if weakly referenced
        receivers are still live. See :meth:`receivers_for` for a stronger
        search.

        :param sender: Check for receivers connected to this sender, in addition
            to those connected to :data:`ANY`.
        """
        if not self.receivers:
            return False

        if self._by_sender[ANY_ID]:
            return True

        if sender is ANY:
            return False

        return make_id(sender) in self._by_sender

    def receivers_for(
        self, sender: t.Any
    ) -> c.Generator[c.Callable[..., t.Any], None, None]:
        """Yield each receiver to be called for ``sender``, in addition to those
        to be called for :data:`ANY`. Weakly referenced receivers that are not
        live will be disconnected and skipped.

        :param sender: Yield receivers connected to this sender, in addition
            to those connected to :data:`ANY`.
        """
        # TODO: test receivers_for(ANY)
        if not self.receivers:
            return

        sender_id = make_id(sender)

        if sender_id in self._by_sender:
            ids = self._by_sender[ANY_ID] | self._by_sender[sender_id]
        else:
            ids = self._by_sender[ANY_ID].copy()

        for receiver_id in ids:
            receiver = self.receivers.get(receiver_id)

            if receiver is None:
                continue

            if isinstance(receiver, weakref.ref):
                strong = receiver()

                if strong is None:
                    self._disconnect(receiver_id, ANY_ID)
                    continue

                yield strong
            else:
                yield receiver

    def disconnect(self, receiver: c.Callable[..., t.Any], sender: t.Any = ANY) -> None:
        """Disconnect ``receiver`` from being called when the signal is sent by
        ``sender``.

        :param receiver: A connected receiver callable.
        :param sender: Disconnect from only this sender. By default, disconnect
            from all senders.
        """
        sender_id: c.Hashable

        if sender is ANY:
            sender_id = ANY_ID
        else:
            sender_id = make_id(sender)

        receiver_id = make_id(receiver)
        self._disconnect(receiver_id, sender_id)

        if (
            "receiver_disconnected" in self.__dict__
            and self.receiver_disconnected.receivers
        ):
            self.receiver_disconnected.send(self, receiver=receiver, sender=sender)

    def _disconnect(self, receiver_id: c.Hashable, sender_id: c.Hashable) -> None:
        if sender_id == ANY_ID:
            if self._by_receiver.pop(receiver_id, None) is not None:
                for bucket in self._by_sender.values():
                    bucket.discard(receiver_id)

            self.receivers.pop(receiver_id, None)
        else:
            self._by_sender[sender_id].discard(receiver_id)
            self._by_receiver[receiver_id].discard(sender_id)

    def _make_cleanup_receiver(
        self, receiver_id: c.Hashable
    ) -> c.Callable[[weakref.ref[c.Callable[..., t.Any]]], None]:
        """Create a callback function to disconnect a weakly referenced
        receiver when it is garbage collected.
        """

        def cleanup(ref: weakref.ref[c.Callable[..., t.Any]]) -> None:
            # If the interpreter is shutting down, disconnecting can result in a
            # weird ignored exception. Don't call it in that case.
            if not sys.is_finalizing():
                self._disconnect(receiver_id, ANY_ID)

        return cleanup

    def _make_cleanup_sender(
        self, sender_id: c.Hashable
    ) -> c.Callable[[weakref.ref[t.Any]], None]:
        """Create a callback function to disconnect all receivers for a weakly
        referenced sender when it is garbage collected.
        """
        assert sender_id != ANY_ID

        def cleanup(ref: weakref.ref[t.Any]) -> None:
            self._weak_senders.pop(sender_id, None)

            for receiver_id in self._by_sender.pop(sender_id, ()):
                self._by_receiver[receiver_id].discard(sender_id)

        return cleanup

    def _cleanup_bookkeeping(self) -> None:
        """Prune unused sender/receiver bookkeeping. Not threadsafe.

        Connecting & disconnecting leaves behind a small amount of bookkeeping
        data. Typical workloads using Blinker, for example in most web apps,
        Flask, CLI scripts, etc., are not adversely affected by this
        bookkeeping.

        With a long-running process performing dynamic signal routing with high
        volume, e.g. connecting to function closures, senders are all unique
        object instances. Doing all of this over and over may cause memory usage
        to grow due to extraneous bookkeeping. (An empty ``set`` for each stale
        sender/receiver pair.)

        This method will prune that bookkeeping away, with the caveat that such
        pruning is not threadsafe. The risk is that cleanup of a fully
        disconnected receiver/sender pair occurs while another thread is
        connecting that same pair. If you are in the highly dynamic, unique
        receiver/sender situation that has lead you to this method, that failure
        mode is perhaps not a big deal for you.
        """
        for mapping in (self._by_sender, self._by_receiver):
            for ident, bucket in list(mapping.items()):
                if not bucket:
                    mapping.pop(ident, None)

    def _clear_state(self) -> None:
        """Disconnect all receivers and senders. Useful for tests."""
        self._weak_senders.clear()
        self.receivers.clear()
        self._by_sender.clear()
        self._by_receiver.clear()


class NamedSignal(Signal):
    """A named generic notification emitter. The name is not used by the signal
    itself, but matches the key in the :class:`Namespace` that it belongs to.

    :param name: The name of the signal within the namespace.
    :param doc: The docstring for the signal.
    """

    def __init__(self, name: str, doc: str | None = None) -> None:
        super().__init__(doc)

        #: The name of this signal.
        self.name: str = name

    def __repr__(self) -> str:
        base = super().__repr__()
        return f"{base[:-1]}; {self.name!r}>"  # noqa: E702


class Namespace(dict[str, NamedSignal]):
    """A dict mapping names to signals."""

    def signal(self, name: str, doc: str | None = None) -> NamedSignal:
        """Return the :class:`NamedSignal` for the given ``name``, creating it
        if required. Repeated calls with the same name return the same signal.

        :param name: The name of the signal.
        :param doc: The docstring of the signal.
        """
        if name not in self:
            self[name] = NamedSignal(name, doc)

        return self[name]


class _PNamespaceSignal(t.Protocol):
    def __call__(self, name: str, doc: str | None = None) -> NamedSignal: ...


default_namespace: Namespace = Namespace()
"""A default :class:`Namespace` for creating named signals. :func:`signal`
creates a :class:`NamedSignal` in this namespace.
"""

signal: _PNamespaceSignal = default_namespace.signal
"""Return a :class:`NamedSignal` in :data:`default_namespace` with the given
``name``, creating it if required. Repeated calls with the same name return the
same signal.
"""

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\worksheet\cell_range.py ===
# Copyright (c) 2010-2024 openpyxl

from copy import copy
from operator import attrgetter

from openpyxl.descriptors import Strict
from openpyxl.descriptors import MinMax
from openpyxl.descriptors.sequence import UniqueSequence
from openpyxl.descriptors.serialisable import Serialisable

from openpyxl.utils import (
    range_boundaries,
    range_to_tuple,
    get_column_letter,
    quote_sheetname,
)

class CellRange(Serialisable):
    """
    Represents a range in a sheet: title and coordinates.

    This object is used to perform operations on ranges, like:

    - shift, expand or shrink
    - union/intersection with another sheet range,

    We can check whether a range is:

    - equal or not equal to another,
    - disjoint of another,
    - contained in another.

    We can get:

    - the size of a range.
    - the range bounds (vertices)
    - the coordinates,
    - the string representation,

    """

    min_col = MinMax(min=1, max=18278, expected_type=int)
    min_row = MinMax(min=1, max=1048576, expected_type=int)
    max_col = MinMax(min=1, max=18278, expected_type=int)
    max_row = MinMax(min=1, max=1048576, expected_type=int)


    def __init__(self, range_string=None, min_col=None, min_row=None,
                 max_col=None, max_row=None, title=None):
        if range_string is not None:
            if "!" in range_string:
                title, (min_col, min_row, max_col, max_row) = range_to_tuple(range_string)
            else:
                min_col, min_row, max_col, max_row = range_boundaries(range_string)

        self.min_col = min_col
        self.min_row = min_row
        self.max_col = max_col
        self.max_row = max_row
        self.title = title

        if min_col > max_col:
            fmt = "{max_col} must be greater than {min_col}"
            raise ValueError(fmt.format(min_col=min_col, max_col=max_col))
        if min_row > max_row:
            fmt = "{max_row} must be greater than {min_row}"
            raise ValueError(fmt.format(min_row=min_row, max_row=max_row))


    @property
    def bounds(self):
        """
        Vertices of the range as a tuple
        """
        return self.min_col, self.min_row, self.max_col, self.max_row


    @property
    def coord(self):
        """
        Excel-style representation of the range
        """
        fmt = "{min_col}{min_row}:{max_col}{max_row}"
        if (self.min_col == self.max_col
            and self.min_row == self.max_row):
            fmt = "{min_col}{min_row}"

        return fmt.format(
            min_col=get_column_letter(self.min_col),
            min_row=self.min_row,
            max_col=get_column_letter(self.max_col),
            max_row=self.max_row
        )

    @property
    def rows(self):
        """
        Return cell coordinates as rows
        """
        for row in range(self.min_row, self.max_row+1):
            yield [(row, col) for col in range(self.min_col, self.max_col+1)]


    @property
    def cols(self):
        """
        Return cell coordinates as columns
        """
        for col in range(self.min_col, self.max_col+1):
            yield [(row, col) for row in range(self.min_row, self.max_row+1)]


    @property
    def cells(self):
        from itertools import product
        return product(range(self.min_row, self.max_row+1), range(self.min_col, self.max_col+1))


    def _check_title(self, other):
        """
        Check whether comparisons between ranges are possible.
        Cannot compare ranges from different worksheets
        Skip if the range passed in has no title.
        """
        if not isinstance(other, CellRange):
            raise TypeError(repr(type(other)))

        if other.title and self.title != other.title:
            raise ValueError("Cannot work with ranges from different worksheets")


    def __repr__(self):
        fmt = u"<{cls} {coord}>"
        if self.title:
            fmt = u"<{cls} {title!r}!{coord}>"
        return fmt.format(cls=self.__class__.__name__, title=self.title, coord=self.coord)


    def __hash__(self):
        return hash((self.min_row, self.min_col, self.max_row, self.max_col))


    def __str__(self):
        fmt = "{coord}"
        title = self.title
        if title:
            fmt = u"{title}!{coord}"
            title = quote_sheetname(title)
        return fmt.format(title=title, coord=self.coord)


    def __copy__(self):
        return self.__class__(min_col=self.min_col, min_row=self.min_row,
                              max_col=self.max_col, max_row=self.max_row,
                              title=self.title)


    def shift(self, col_shift=0, row_shift=0):
        """
        Shift the focus of the range according to the shift values (*col_shift*, *row_shift*).

        :type col_shift: int
        :param col_shift: number of columns to be moved by, can be negative
        :type row_shift: int
        :param row_shift: number of rows to be moved by, can be negative
        :raise: :class:`ValueError` if any row or column index < 1
        """

        if (self.min_col + col_shift <= 0
            or self.min_row + row_shift <= 0):
            raise ValueError("Invalid shift value: col_shift={0}, row_shift={1}".format(col_shift, row_shift))
        self.min_col += col_shift
        self.min_row += row_shift
        self.max_col += col_shift
        self.max_row += row_shift


    def __ne__(self, other):
        """
        Test whether the ranges are not equal.

        :type other: openpyxl.worksheet.cell_range.CellRange
        :param other: Other sheet range
        :return: ``True`` if *range* != *other*.
        """
        try:
            self._check_title(other)
        except ValueError:
            return True

        return (
            other.min_row != self.min_row
            or self.max_row != other.max_row
            or other.min_col != self.min_col
            or self.max_col != other.max_col
        )


    def __eq__(self, other):
        """
        Test whether the ranges are equal.

        :type other: openpyxl.worksheet.cell_range.CellRange
        :param other: Other sheet range
        :return: ``True`` if *range* == *other*.
        """
        return not self.__ne__(other)


    def issubset(self, other):
        """
        Test whether every cell in this range is also in *other*.

        :type other: openpyxl.worksheet.cell_range.CellRange
        :param other: Other sheet range
        :return: ``True`` if *range* <= *other*.
        """
        self._check_title(other)

        return other.__superset(self)

    __le__ = issubset


    def __lt__(self, other):
        """
        Test whether *other* contains every cell of this range, and more.

        :type other: openpyxl.worksheet.cell_range.CellRange
        :param other: Other sheet range
        :return: ``True`` if *range* < *other*.
        """
        return self.__le__(other) and self.__ne__(other)


    def __superset(self, other):
        return (
            (self.min_row <= other.min_row <= other.max_row <= self.max_row)
            and
            (self.min_col <= other.min_col <= other.max_col <= self.max_col)
        )


    def issuperset(self, other):
        """
        Test whether every cell in *other* is in this range.

        :type other: openpyxl.worksheet.cell_range.CellRange
        :param other: Other sheet range
        :return: ``True`` if *range* >= *other* (or *other* in *range*).
        """
        self._check_title(other)

        return self.__superset(other)

    __ge__ = issuperset


    def __contains__(self, coord):
        """
        Check whether the range contains a particular cell coordinate
        """
        cr = self.__class__(coord)
        return self.__superset(cr)


    def __gt__(self, other):
        """
        Test whether this range contains every cell in *other*, and more.

        :type other: openpyxl.worksheet.cell_range.CellRange
        :param other: Other sheet range
        :return: ``True`` if *range* > *other*.
        """
        return self.__ge__(other) and self.__ne__(other)


    def isdisjoint(self, other):
        """
        Return ``True`` if this range has no cell in common with *other*.
        Ranges are disjoint if and only if their intersection is the empty range.

        :type other: openpyxl.worksheet.cell_range.CellRange
        :param other: Other sheet range.
        :return: ``True`` if the range has no cells in common with other.
        """
        self._check_title(other)

        # Sort by top-left vertex
        if self.bounds > other.bounds:
            self, other = other, self

        return (self.max_col < other.min_col
                or self.max_row < other.min_row
                or other.max_row < self.min_row)


    def intersection(self, other):
        """
        Return a new range with cells common to this range and *other*

        :type other: openpyxl.worksheet.cell_range.CellRange
        :param other: Other sheet range.
        :return: the intersecting sheet range.
        :raise: :class:`ValueError` if the *other* range doesn't intersect
            with this range.
        """
        if self.isdisjoint(other):
            raise ValueError("Range {0} doesn't intersect {0}".format(self, other))

        min_row = max(self.min_row, other.min_row)
        max_row = min(self.max_row, other.max_row)
        min_col = max(self.min_col, other.min_col)
        max_col = min(self.max_col, other.max_col)

        return CellRange(min_col=min_col, min_row=min_row, max_col=max_col,
                         max_row=max_row)

    __and__ = intersection


    def union(self, other):
        """
        Return the minimal superset of this range and *other*. This new range
        will contain all cells from this range, *other*, and any additional
        cells required to form a rectangular ``CellRange``.

        :type other: openpyxl.worksheet.cell_range.CellRange
        :param other: Other sheet range.
        :return: a ``CellRange`` that is a superset of this and *other*.
        """
        self._check_title(other)

        min_row = min(self.min_row, other.min_row)
        max_row = max(self.max_row, other.max_row)
        min_col = min(self.min_col, other.min_col)
        max_col = max(self.max_col, other.max_col)
        return CellRange(min_col=min_col, min_row=min_row, max_col=max_col,
                         max_row=max_row, title=self.title)

    __or__ = union


    def __iter__(self):
        """
        For use as a dictionary elsewhere in the library.
        """
        for x in self.__attrs__:
            if x == "title":
                continue
            v = getattr(self, x)
            yield x, v


    def expand(self, right=0, down=0, left=0, up=0):
        """
        Expand the range by the dimensions provided.

        :type right: int
        :param right: expand range to the right by this number of cells
        :type down: int
        :param down: expand range down by this number of cells
        :type left: int
        :param left: expand range to the left by this number of cells
        :type up: int
        :param up: expand range up by this number of cells
        """
        self.min_col -= left
        self.min_row -= up
        self.max_col += right
        self.max_row += down


    def shrink(self, right=0, bottom=0, left=0, top=0):
        """
        Shrink the range by the dimensions provided.

        :type right: int
        :param right: shrink range from the right by this number of cells
        :type down: int
        :param down: shrink range from the top by this number of cells
        :type left: int
        :param left: shrink range from the left by this number of cells
        :type up: int
        :param up: shrink range from the bottom by this number of cells
        """
        self.min_col += left
        self.min_row += top
        self.max_col -= right
        self.max_row -= bottom


    @property
    def size(self):
        """ Return the size of the range as a dictionary of rows and columns. """
        cols = self.max_col + 1 - self.min_col
        rows = self.max_row + 1 - self.min_row
        return {'columns':cols, 'rows':rows}


    @property
    def top(self):
        """A list of cell coordinates that comprise the top of the range"""
        return [(self.min_row, col) for col in range(self.min_col, self.max_col+1)]


    @property
    def bottom(self):
        """A list of cell coordinates that comprise the bottom of the range"""
        return [(self.max_row, col) for col in range(self.min_col, self.max_col+1)]


    @property
    def left(self):
        """A list of cell coordinates that comprise the left-side of the range"""
        return [(row, self.min_col) for row in range(self.min_row, self.max_row+1)]


    @property
    def right(self):
        """A list of cell coordinates that comprise the right-side of the range"""
        return [(row, self.max_col) for row in range(self.min_row, self.max_row+1)]


class MultiCellRange(Strict):


    ranges = UniqueSequence(expected_type=CellRange)


    def __init__(self, ranges=set()):
        if isinstance(ranges, str):
            ranges = [CellRange(r) for r in ranges.split()]
        self.ranges = set(ranges)


    def __contains__(self, coord):
        if isinstance(coord, str):
            coord = CellRange(coord)
        for r in self.ranges:
            if coord <= r:
                return True
        return False


    def __repr__(self):
        ranges = " ".join([str(r) for r in self.sorted()])
        return f"<{self.__class__.__name__} [{ranges}]>"


    def __str__(self):
        ranges = u" ".join([str(r) for r in self.sorted()])
        return ranges


    def __hash__(self):
        return hash(str(self))


    def sorted(self):
        """
        Return a sorted list of items
        """
        return sorted(self.ranges, key=attrgetter('min_col', 'min_row', 'max_col', 'max_row'))


    def add(self, coord):
        """
        Add a cell coordinate or CellRange
        """
        cr = coord
        if isinstance(coord, str):
            cr = CellRange(coord)
        elif not isinstance(coord, CellRange):
            raise ValueError("You can only add CellRanges")
        if cr not in self:
            self.ranges.add(cr)


    def __iadd__(self, coord):
        self.add(coord)
        return self


    def __eq__(self, other):
        if  isinstance(other, str):
            other = self.__class__(other)
        return self.ranges == other.ranges


    def __ne__(self, other):
        return not self == other


    def __bool__(self):
        return bool(self.ranges)


    def remove(self, coord):
        if not isinstance(coord, CellRange):
            coord = CellRange(coord)
        self.ranges.remove(coord)


    def __iter__(self):
        for cr in self.ranges:
            yield cr


    def __copy__(self):
        ranges = {copy(r) for r in self.ranges}
        return MultiCellRange(ranges)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\reshape\melt.py ===
from __future__ import annotations

import re
from typing import TYPE_CHECKING

import numpy as np

from pandas.util._decorators import Appender

from pandas.core.dtypes.common import is_list_like
from pandas.core.dtypes.concat import concat_compat
from pandas.core.dtypes.missing import notna

import pandas.core.algorithms as algos
from pandas.core.indexes.api import MultiIndex
from pandas.core.reshape.concat import concat
from pandas.core.reshape.util import tile_compat
from pandas.core.shared_docs import _shared_docs
from pandas.core.tools.numeric import to_numeric

if TYPE_CHECKING:
    from collections.abc import Hashable

    from pandas._typing import AnyArrayLike

    from pandas import DataFrame


def ensure_list_vars(arg_vars, variable: str, columns) -> list:
    if arg_vars is not None:
        if not is_list_like(arg_vars):
            return [arg_vars]
        elif isinstance(columns, MultiIndex) and not isinstance(arg_vars, list):
            raise ValueError(
                f"{variable} must be a list of tuples when columns are a MultiIndex"
            )
        else:
            return list(arg_vars)
    else:
        return []


@Appender(_shared_docs["melt"] % {"caller": "pd.melt(df, ", "other": "DataFrame.melt"})
def melt(
    frame: DataFrame,
    id_vars=None,
    value_vars=None,
    var_name=None,
    value_name: Hashable = "value",
    col_level=None,
    ignore_index: bool = True,
) -> DataFrame:
    if value_name in frame.columns:
        raise ValueError(
            f"value_name ({value_name}) cannot match an element in "
            "the DataFrame columns."
        )
    id_vars = ensure_list_vars(id_vars, "id_vars", frame.columns)
    value_vars_was_not_none = value_vars is not None
    value_vars = ensure_list_vars(value_vars, "value_vars", frame.columns)

    if id_vars or value_vars:
        if col_level is not None:
            level = frame.columns.get_level_values(col_level)
        else:
            level = frame.columns
        labels = id_vars + value_vars
        idx = level.get_indexer_for(labels)
        missing = idx == -1
        if missing.any():
            missing_labels = [
                lab for lab, not_found in zip(labels, missing) if not_found
            ]
            raise KeyError(
                "The following id_vars or value_vars are not present in "
                f"the DataFrame: {missing_labels}"
            )
        if value_vars_was_not_none:
            frame = frame.iloc[:, algos.unique(idx)]
        else:
            frame = frame.copy()
    else:
        frame = frame.copy()

    if col_level is not None:  # allow list or other?
        # frame is a copy
        frame.columns = frame.columns.get_level_values(col_level)

    if var_name is None:
        if isinstance(frame.columns, MultiIndex):
            if len(frame.columns.names) == len(set(frame.columns.names)):
                var_name = frame.columns.names
            else:
                var_name = [f"variable_{i}" for i in range(len(frame.columns.names))]
        else:
            var_name = [
                frame.columns.name if frame.columns.name is not None else "variable"
            ]
    elif is_list_like(var_name):
        raise ValueError(f"{var_name=} must be a scalar.")
    else:
        var_name = [var_name]

    num_rows, K = frame.shape
    num_cols_adjusted = K - len(id_vars)

    mdata: dict[Hashable, AnyArrayLike] = {}
    for col in id_vars:
        id_data = frame.pop(col)
        if not isinstance(id_data.dtype, np.dtype):
            # i.e. ExtensionDtype
            if num_cols_adjusted > 0:
                mdata[col] = concat([id_data] * num_cols_adjusted, ignore_index=True)
            else:
                # We can't concat empty list. (GH 46044)
                mdata[col] = type(id_data)([], name=id_data.name, dtype=id_data.dtype)
        else:
            mdata[col] = np.tile(id_data._values, num_cols_adjusted)

    mcolumns = id_vars + var_name + [value_name]

    if frame.shape[1] > 0 and not any(
        not isinstance(dt, np.dtype) and dt._supports_2d for dt in frame.dtypes
    ):
        mdata[value_name] = concat(
            [frame.iloc[:, i] for i in range(frame.shape[1])]
        ).values
    else:
        mdata[value_name] = frame._values.ravel("F")
    for i, col in enumerate(var_name):
        mdata[col] = frame.columns._get_level_values(i).repeat(num_rows)

    result = frame._constructor(mdata, columns=mcolumns)

    if not ignore_index:
        result.index = tile_compat(frame.index, num_cols_adjusted)

    return result


def lreshape(data: DataFrame, groups: dict, dropna: bool = True) -> DataFrame:
    """
    Reshape wide-format data to long. Generalized inverse of DataFrame.pivot.

    Accepts a dictionary, ``groups``, in which each key is a new column name
    and each value is a list of old column names that will be "melted" under
    the new column name as part of the reshape.

    Parameters
    ----------
    data : DataFrame
        The wide-format DataFrame.
    groups : dict
        {new_name : list_of_columns}.
    dropna : bool, default True
        Do not include columns whose entries are all NaN.

    Returns
    -------
    DataFrame
        Reshaped DataFrame.

    See Also
    --------
    melt : Unpivot a DataFrame from wide to long format, optionally leaving
        identifiers set.
    pivot : Create a spreadsheet-style pivot table as a DataFrame.
    DataFrame.pivot : Pivot without aggregation that can handle
        non-numeric data.
    DataFrame.pivot_table : Generalization of pivot that can handle
        duplicate values for one index/column pair.
    DataFrame.unstack : Pivot based on the index values instead of a
        column.
    wide_to_long : Wide panel to long format. Less flexible but more
        user-friendly than melt.

    Examples
    --------
    >>> data = pd.DataFrame({'hr1': [514, 573], 'hr2': [545, 526],
    ...                      'team': ['Red Sox', 'Yankees'],
    ...                      'year1': [2007, 2007], 'year2': [2008, 2008]})
    >>> data
       hr1  hr2     team  year1  year2
    0  514  545  Red Sox   2007   2008
    1  573  526  Yankees   2007   2008

    >>> pd.lreshape(data, {'year': ['year1', 'year2'], 'hr': ['hr1', 'hr2']})
          team  year   hr
    0  Red Sox  2007  514
    1  Yankees  2007  573
    2  Red Sox  2008  545
    3  Yankees  2008  526
    """
    mdata = {}
    pivot_cols = []
    all_cols: set[Hashable] = set()
    K = len(next(iter(groups.values())))
    for target, names in groups.items():
        if len(names) != K:
            raise ValueError("All column lists must be same length")
        to_concat = [data[col]._values for col in names]

        mdata[target] = concat_compat(to_concat)
        pivot_cols.append(target)
        all_cols = all_cols.union(names)

    id_cols = list(data.columns.difference(all_cols))
    for col in id_cols:
        mdata[col] = np.tile(data[col]._values, K)

    if dropna:
        mask = np.ones(len(mdata[pivot_cols[0]]), dtype=bool)
        for c in pivot_cols:
            mask &= notna(mdata[c])
        if not mask.all():
            mdata = {k: v[mask] for k, v in mdata.items()}

    return data._constructor(mdata, columns=id_cols + pivot_cols)


def wide_to_long(
    df: DataFrame, stubnames, i, j, sep: str = "", suffix: str = r"\d+"
) -> DataFrame:
    r"""
    Unpivot a DataFrame from wide to long format.

    Less flexible but more user-friendly than melt.

    With stubnames ['A', 'B'], this function expects to find one or more
    group of columns with format
    A-suffix1, A-suffix2,..., B-suffix1, B-suffix2,...
    You specify what you want to call this suffix in the resulting long format
    with `j` (for example `j='year'`)

    Each row of these wide variables are assumed to be uniquely identified by
    `i` (can be a single column name or a list of column names)

    All remaining variables in the data frame are left intact.

    Parameters
    ----------
    df : DataFrame
        The wide-format DataFrame.
    stubnames : str or list-like
        The stub name(s). The wide format variables are assumed to
        start with the stub names.
    i : str or list-like
        Column(s) to use as id variable(s).
    j : str
        The name of the sub-observation variable. What you wish to name your
        suffix in the long format.
    sep : str, default ""
        A character indicating the separation of the variable names
        in the wide format, to be stripped from the names in the long format.
        For example, if your column names are A-suffix1, A-suffix2, you
        can strip the hyphen by specifying `sep='-'`.
    suffix : str, default '\\d+'
        A regular expression capturing the wanted suffixes. '\\d+' captures
        numeric suffixes. Suffixes with no numbers could be specified with the
        negated character class '\\D+'. You can also further disambiguate
        suffixes, for example, if your wide variables are of the form A-one,
        B-two,.., and you have an unrelated column A-rating, you can ignore the
        last one by specifying `suffix='(!?one|two)'`. When all suffixes are
        numeric, they are cast to int64/float64.

    Returns
    -------
    DataFrame
        A DataFrame that contains each stub name as a variable, with new index
        (i, j).

    See Also
    --------
    melt : Unpivot a DataFrame from wide to long format, optionally leaving
        identifiers set.
    pivot : Create a spreadsheet-style pivot table as a DataFrame.
    DataFrame.pivot : Pivot without aggregation that can handle
        non-numeric data.
    DataFrame.pivot_table : Generalization of pivot that can handle
        duplicate values for one index/column pair.
    DataFrame.unstack : Pivot based on the index values instead of a
        column.

    Notes
    -----
    All extra variables are left untouched. This simply uses
    `pandas.melt` under the hood, but is hard-coded to "do the right thing"
    in a typical case.

    Examples
    --------
    >>> np.random.seed(123)
    >>> df = pd.DataFrame({"A1970" : {0 : "a", 1 : "b", 2 : "c"},
    ...                    "A1980" : {0 : "d", 1 : "e", 2 : "f"},
    ...                    "B1970" : {0 : 2.5, 1 : 1.2, 2 : .7},
    ...                    "B1980" : {0 : 3.2, 1 : 1.3, 2 : .1},
    ...                    "X"     : dict(zip(range(3), np.random.randn(3)))
    ...                   })
    >>> df["id"] = df.index
    >>> df
      A1970 A1980  B1970  B1980         X  id
    0     a     d    2.5    3.2 -1.085631   0
    1     b     e    1.2    1.3  0.997345   1
    2     c     f    0.7    0.1  0.282978   2
    >>> pd.wide_to_long(df, ["A", "B"], i="id", j="year")
    ... # doctest: +NORMALIZE_WHITESPACE
                    X  A    B
    id year
    0  1970 -1.085631  a  2.5
    1  1970  0.997345  b  1.2
    2  1970  0.282978  c  0.7
    0  1980 -1.085631  d  3.2
    1  1980  0.997345  e  1.3
    2  1980  0.282978  f  0.1

    With multiple id columns

    >>> df = pd.DataFrame({
    ...     'famid': [1, 1, 1, 2, 2, 2, 3, 3, 3],
    ...     'birth': [1, 2, 3, 1, 2, 3, 1, 2, 3],
    ...     'ht1': [2.8, 2.9, 2.2, 2, 1.8, 1.9, 2.2, 2.3, 2.1],
    ...     'ht2': [3.4, 3.8, 2.9, 3.2, 2.8, 2.4, 3.3, 3.4, 2.9]
    ... })
    >>> df
       famid  birth  ht1  ht2
    0      1      1  2.8  3.4
    1      1      2  2.9  3.8
    2      1      3  2.2  2.9
    3      2      1  2.0  3.2
    4      2      2  1.8  2.8
    5      2      3  1.9  2.4
    6      3      1  2.2  3.3
    7      3      2  2.3  3.4
    8      3      3  2.1  2.9
    >>> l = pd.wide_to_long(df, stubnames='ht', i=['famid', 'birth'], j='age')
    >>> l
    ... # doctest: +NORMALIZE_WHITESPACE
                      ht
    famid birth age
    1     1     1    2.8
                2    3.4
          2     1    2.9
                2    3.8
          3     1    2.2
                2    2.9
    2     1     1    2.0
                2    3.2
          2     1    1.8
                2    2.8
          3     1    1.9
                2    2.4
    3     1     1    2.2
                2    3.3
          2     1    2.3
                2    3.4
          3     1    2.1
                2    2.9

    Going from long back to wide just takes some creative use of `unstack`

    >>> w = l.unstack()
    >>> w.columns = w.columns.map('{0[0]}{0[1]}'.format)
    >>> w.reset_index()
       famid  birth  ht1  ht2
    0      1      1  2.8  3.4
    1      1      2  2.9  3.8
    2      1      3  2.2  2.9
    3      2      1  2.0  3.2
    4      2      2  1.8  2.8
    5      2      3  1.9  2.4
    6      3      1  2.2  3.3
    7      3      2  2.3  3.4
    8      3      3  2.1  2.9

    Less wieldy column names are also handled

    >>> np.random.seed(0)
    >>> df = pd.DataFrame({'A(weekly)-2010': np.random.rand(3),
    ...                    'A(weekly)-2011': np.random.rand(3),
    ...                    'B(weekly)-2010': np.random.rand(3),
    ...                    'B(weekly)-2011': np.random.rand(3),
    ...                    'X' : np.random.randint(3, size=3)})
    >>> df['id'] = df.index
    >>> df # doctest: +NORMALIZE_WHITESPACE, +ELLIPSIS
       A(weekly)-2010  A(weekly)-2011  B(weekly)-2010  B(weekly)-2011  X  id
    0        0.548814        0.544883        0.437587        0.383442  0   0
    1        0.715189        0.423655        0.891773        0.791725  1   1
    2        0.602763        0.645894        0.963663        0.528895  1   2

    >>> pd.wide_to_long(df, ['A(weekly)', 'B(weekly)'], i='id',
    ...                 j='year', sep='-')
    ... # doctest: +NORMALIZE_WHITESPACE
             X  A(weekly)  B(weekly)
    id year
    0  2010  0   0.548814   0.437587
    1  2010  1   0.715189   0.891773
    2  2010  1   0.602763   0.963663
    0  2011  0   0.544883   0.383442
    1  2011  1   0.423655   0.791725
    2  2011  1   0.645894   0.528895

    If we have many columns, we could also use a regex to find our
    stubnames and pass that list on to wide_to_long

    >>> stubnames = sorted(
    ...     set([match[0] for match in df.columns.str.findall(
    ...         r'[A-B]\(.*\)').values if match != []])
    ... )
    >>> list(stubnames)
    ['A(weekly)', 'B(weekly)']

    All of the above examples have integers as suffixes. It is possible to
    have non-integers as suffixes.

    >>> df = pd.DataFrame({
    ...     'famid': [1, 1, 1, 2, 2, 2, 3, 3, 3],
    ...     'birth': [1, 2, 3, 1, 2, 3, 1, 2, 3],
    ...     'ht_one': [2.8, 2.9, 2.2, 2, 1.8, 1.9, 2.2, 2.3, 2.1],
    ...     'ht_two': [3.4, 3.8, 2.9, 3.2, 2.8, 2.4, 3.3, 3.4, 2.9]
    ... })
    >>> df
       famid  birth  ht_one  ht_two
    0      1      1     2.8     3.4
    1      1      2     2.9     3.8
    2      1      3     2.2     2.9
    3      2      1     2.0     3.2
    4      2      2     1.8     2.8
    5      2      3     1.9     2.4
    6      3      1     2.2     3.3
    7      3      2     2.3     3.4
    8      3      3     2.1     2.9

    >>> l = pd.wide_to_long(df, stubnames='ht', i=['famid', 'birth'], j='age',
    ...                     sep='_', suffix=r'\w+')
    >>> l
    ... # doctest: +NORMALIZE_WHITESPACE
                      ht
    famid birth age
    1     1     one  2.8
                two  3.4
          2     one  2.9
                two  3.8
          3     one  2.2
                two  2.9
    2     1     one  2.0
                two  3.2
          2     one  1.8
                two  2.8
          3     one  1.9
                two  2.4
    3     1     one  2.2
                two  3.3
          2     one  2.3
                two  3.4
          3     one  2.1
                two  2.9
    """

    def get_var_names(df, stub: str, sep: str, suffix: str):
        regex = rf"^{re.escape(stub)}{re.escape(sep)}{suffix}$"
        return df.columns[df.columns.str.match(regex)]

    def melt_stub(df, stub: str, i, j, value_vars, sep: str):
        newdf = melt(
            df,
            id_vars=i,
            value_vars=value_vars,
            value_name=stub.rstrip(sep),
            var_name=j,
        )
        newdf[j] = newdf[j].str.replace(re.escape(stub + sep), "", regex=True)

        # GH17627 Cast numerics suffixes to int/float
        try:
            newdf[j] = to_numeric(newdf[j])
        except (TypeError, ValueError, OverflowError):
            # TODO: anything else to catch?
            pass

        return newdf.set_index(i + [j])

    if not is_list_like(stubnames):
        stubnames = [stubnames]
    else:
        stubnames = list(stubnames)

    if df.columns.isin(stubnames).any():
        raise ValueError("stubname can't be identical to a column name")

    if not is_list_like(i):
        i = [i]
    else:
        i = list(i)

    if df[i].duplicated().any():
        raise ValueError("the id variables need to uniquely identify each row")

    _melted = []
    value_vars_flattened = []
    for stub in stubnames:
        value_var = get_var_names(df, stub, sep, suffix)
        value_vars_flattened.extend(value_var)
        _melted.append(melt_stub(df, stub, i, j, value_var, sep))

    melted = concat(_melted, axis=1)
    id_vars = df.columns.difference(value_vars_flattened)
    new = df[id_vars]

    if len(i) == 1:
        return new.set_index(i).join(melted)
    else:
        return new.merge(melted.reset_index(), on=i).set_index(i + [j])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\parsing\latex\_antlr\latexlexer.py ===
# *** GENERATED BY `setup.py antlr`, DO NOT EDIT BY HAND ***
#
# Generated from ../LaTeX.g4, derived from latex2sympy
#     latex2sympy is licensed under the MIT license
#     https://github.com/augustt198/latex2sympy/blob/master/LICENSE.txt
#
# Generated with antlr4
#    antlr4 is licensed under the BSD-3-Clause License
#    https://github.com/antlr/antlr4/blob/master/LICENSE.txt
from antlr4 import *
from io import StringIO
import sys
if sys.version_info[1] > 5:
    from typing import TextIO
else:
    from typing.io import TextIO


def serializedATN():
    return [
        4,0,91,911,6,-1,2,0,7,0,2,1,7,1,2,2,7,2,2,3,7,3,2,4,7,4,2,5,7,5,
        2,6,7,6,2,7,7,7,2,8,7,8,2,9,7,9,2,10,7,10,2,11,7,11,2,12,7,12,2,
        13,7,13,2,14,7,14,2,15,7,15,2,16,7,16,2,17,7,17,2,18,7,18,2,19,7,
        19,2,20,7,20,2,21,7,21,2,22,7,22,2,23,7,23,2,24,7,24,2,25,7,25,2,
        26,7,26,2,27,7,27,2,28,7,28,2,29,7,29,2,30,7,30,2,31,7,31,2,32,7,
        32,2,33,7,33,2,34,7,34,2,35,7,35,2,36,7,36,2,37,7,37,2,38,7,38,2,
        39,7,39,2,40,7,40,2,41,7,41,2,42,7,42,2,43,7,43,2,44,7,44,2,45,7,
        45,2,46,7,46,2,47,7,47,2,48,7,48,2,49,7,49,2,50,7,50,2,51,7,51,2,
        52,7,52,2,53,7,53,2,54,7,54,2,55,7,55,2,56,7,56,2,57,7,57,2,58,7,
        58,2,59,7,59,2,60,7,60,2,61,7,61,2,62,7,62,2,63,7,63,2,64,7,64,2,
        65,7,65,2,66,7,66,2,67,7,67,2,68,7,68,2,69,7,69,2,70,7,70,2,71,7,
        71,2,72,7,72,2,73,7,73,2,74,7,74,2,75,7,75,2,76,7,76,2,77,7,77,2,
        78,7,78,2,79,7,79,2,80,7,80,2,81,7,81,2,82,7,82,2,83,7,83,2,84,7,
        84,2,85,7,85,2,86,7,86,2,87,7,87,2,88,7,88,2,89,7,89,2,90,7,90,2,
        91,7,91,1,0,1,0,1,1,1,1,1,2,4,2,191,8,2,11,2,12,2,192,1,2,1,2,1,
        3,1,3,1,3,1,3,1,3,1,3,1,3,1,3,1,3,1,3,1,3,1,3,3,3,209,8,3,1,3,1,
        3,1,4,1,4,1,4,1,4,1,4,1,4,1,4,1,4,1,4,1,4,1,4,3,4,224,8,4,1,4,1,
        4,1,5,1,5,1,5,1,5,1,5,1,5,1,5,1,5,1,5,1,5,1,5,1,5,1,5,3,5,241,8,
        5,1,5,1,5,1,6,1,6,1,6,1,6,1,6,1,6,1,6,1,6,1,7,1,7,1,7,1,7,1,7,1,
        7,1,7,1,7,1,7,1,8,1,8,1,8,1,8,1,8,1,8,1,8,1,8,1,8,1,8,1,8,1,8,1,
        8,1,8,1,8,3,8,277,8,8,1,8,1,8,1,9,1,9,1,9,1,9,1,9,1,9,1,9,1,9,1,
        9,1,9,1,9,1,9,1,9,1,9,1,9,1,10,1,10,1,10,1,10,1,10,1,10,1,10,1,10,
        1,10,1,10,1,10,1,10,1,10,1,10,1,10,1,10,1,10,1,11,1,11,1,11,1,11,
        1,11,1,11,1,11,1,11,1,12,1,12,1,12,1,12,1,12,1,12,1,12,1,12,1,12,
        1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,
        1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,
        1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,
        1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,1,13,3,13,
        381,8,13,1,13,1,13,1,14,1,14,1,15,1,15,1,16,1,16,1,17,1,17,1,18,
        1,18,1,19,1,19,1,20,1,20,1,21,1,21,1,22,1,22,1,22,1,23,1,23,1,23,
        1,24,1,24,1,25,1,25,1,26,1,26,1,27,1,27,1,27,1,27,1,27,1,27,1,27,
        1,27,1,28,1,28,1,28,1,28,1,28,1,28,1,28,1,29,1,29,1,29,1,29,1,29,
        1,29,1,29,1,29,1,30,1,30,1,30,1,30,1,30,1,30,1,30,1,30,1,31,1,31,
        1,31,1,31,1,31,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,
        1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,
        1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,
        1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,1,32,
        1,32,1,32,1,32,1,32,1,32,1,32,3,32,504,8,32,1,33,1,33,1,33,1,33,
        1,33,1,33,1,33,1,33,1,33,1,33,1,33,1,33,1,33,1,33,1,33,3,33,521,
        8,33,1,34,1,34,1,34,1,34,1,34,1,35,1,35,1,35,1,35,1,35,1,35,1,36,
        1,36,1,36,1,36,1,36,1,37,1,37,1,37,1,37,1,37,1,38,1,38,1,38,1,38,
        1,39,1,39,1,39,1,39,1,40,1,40,1,40,1,40,1,40,1,41,1,41,1,41,1,41,
        1,41,1,42,1,42,1,42,1,42,1,42,1,43,1,43,1,43,1,43,1,43,1,44,1,44,
        1,44,1,44,1,44,1,45,1,45,1,45,1,45,1,45,1,46,1,46,1,46,1,46,1,46,
        1,46,1,46,1,46,1,47,1,47,1,47,1,47,1,47,1,47,1,47,1,47,1,48,1,48,
        1,48,1,48,1,48,1,48,1,48,1,48,1,49,1,49,1,49,1,49,1,49,1,49,1,49,
        1,49,1,50,1,50,1,50,1,50,1,50,1,50,1,50,1,50,1,51,1,51,1,51,1,51,
        1,51,1,51,1,51,1,51,1,52,1,52,1,52,1,52,1,52,1,52,1,53,1,53,1,53,
        1,53,1,53,1,53,1,54,1,54,1,54,1,54,1,54,1,54,1,55,1,55,1,55,1,55,
        1,55,1,55,1,55,1,55,1,56,1,56,1,56,1,56,1,56,1,56,1,56,1,56,1,57,
        1,57,1,57,1,57,1,57,1,57,1,57,1,57,1,58,1,58,1,58,1,58,1,58,1,58,
        1,58,1,58,1,59,1,59,1,59,1,59,1,59,1,59,1,59,1,59,1,60,1,60,1,60,
        1,60,1,60,1,60,1,60,1,61,1,61,1,61,1,61,1,61,1,61,1,61,1,62,1,62,
        1,62,1,62,1,62,1,62,1,63,1,63,1,63,1,63,1,63,1,63,1,63,1,63,1,63,
        1,63,1,64,1,64,1,64,1,64,1,64,1,64,1,64,1,65,1,65,1,65,1,65,1,65,
        1,65,1,66,1,66,1,66,1,66,1,66,1,67,1,67,1,67,1,67,1,67,1,67,1,67,
        1,67,1,67,1,67,1,67,1,67,1,67,1,67,1,67,1,67,1,67,3,67,753,8,67,
        1,68,1,68,1,68,1,68,1,68,1,68,1,68,1,69,1,69,1,69,1,69,1,69,1,69,
        1,69,1,69,1,70,1,70,1,70,1,70,1,70,1,70,1,70,1,70,1,71,1,71,1,71,
        1,71,1,71,1,71,1,71,1,71,1,72,1,72,1,73,1,73,1,74,1,74,1,75,1,75,
        1,76,1,76,5,76,796,8,76,10,76,12,76,799,9,76,1,76,1,76,1,76,4,76,
        804,8,76,11,76,12,76,805,3,76,808,8,76,1,77,1,77,1,78,1,78,1,79,
        1,79,5,79,816,8,79,10,79,12,79,819,9,79,3,79,821,8,79,1,79,1,79,
        1,79,5,79,826,8,79,10,79,12,79,829,9,79,1,79,3,79,832,8,79,3,79,
        834,8,79,1,80,1,80,1,80,1,80,1,80,1,81,1,81,1,82,1,82,1,82,1,82,
        1,82,1,82,1,82,1,82,1,82,3,82,852,8,82,1,83,1,83,1,83,1,83,1,83,
        1,83,1,84,1,84,1,84,1,84,1,84,1,84,1,84,1,84,1,84,1,84,1,85,1,85,
        1,86,1,86,1,86,1,86,1,86,1,86,1,86,1,86,1,86,3,86,881,8,86,1,87,
        1,87,1,87,1,87,1,87,1,87,1,88,1,88,1,88,1,88,1,88,1,88,1,88,1,88,
        1,88,1,88,1,89,1,89,1,90,4,90,902,8,90,11,90,12,90,903,1,91,1,91,
        4,91,908,8,91,11,91,12,91,909,3,797,817,827,0,92,1,1,3,2,5,3,7,4,
        9,5,11,6,13,7,15,8,17,9,19,10,21,11,23,12,25,13,27,14,29,15,31,16,
        33,17,35,18,37,19,39,20,41,21,43,22,45,23,47,24,49,25,51,26,53,27,
        55,28,57,29,59,30,61,31,63,32,65,33,67,34,69,35,71,36,73,37,75,38,
        77,39,79,40,81,41,83,42,85,43,87,44,89,45,91,46,93,47,95,48,97,49,
        99,50,101,51,103,52,105,53,107,54,109,55,111,56,113,57,115,58,117,
        59,119,60,121,61,123,62,125,63,127,64,129,65,131,66,133,67,135,68,
        137,69,139,70,141,71,143,72,145,73,147,74,149,75,151,0,153,76,155,
        77,157,78,159,79,161,80,163,81,165,82,167,83,169,84,171,85,173,86,
        175,87,177,88,179,89,181,90,183,91,1,0,3,3,0,9,10,13,13,32,32,2,
        0,65,90,97,122,1,0,48,57,949,0,1,1,0,0,0,0,3,1,0,0,0,0,5,1,0,0,0,
        0,7,1,0,0,0,0,9,1,0,0,0,0,11,1,0,0,0,0,13,1,0,0,0,0,15,1,0,0,0,0,
        17,1,0,0,0,0,19,1,0,0,0,0,21,1,0,0,0,0,23,1,0,0,0,0,25,1,0,0,0,0,
        27,1,0,0,0,0,29,1,0,0,0,0,31,1,0,0,0,0,33,1,0,0,0,0,35,1,0,0,0,0,
        37,1,0,0,0,0,39,1,0,0,0,0,41,1,0,0,0,0,43,1,0,0,0,0,45,1,0,0,0,0,
        47,1,0,0,0,0,49,1,0,0,0,0,51,1,0,0,0,0,53,1,0,0,0,0,55,1,0,0,0,0,
        57,1,0,0,0,0,59,1,0,0,0,0,61,1,0,0,0,0,63,1,0,0,0,0,65,1,0,0,0,0,
        67,1,0,0,0,0,69,1,0,0,0,0,71,1,0,0,0,0,73,1,0,0,0,0,75,1,0,0,0,0,
        77,1,0,0,0,0,79,1,0,0,0,0,81,1,0,0,0,0,83,1,0,0,0,0,85,1,0,0,0,0,
        87,1,0,0,0,0,89,1,0,0,0,0,91,1,0,0,0,0,93,1,0,0,0,0,95,1,0,0,0,0,
        97,1,0,0,0,0,99,1,0,0,0,0,101,1,0,0,0,0,103,1,0,0,0,0,105,1,0,0,
        0,0,107,1,0,0,0,0,109,1,0,0,0,0,111,1,0,0,0,0,113,1,0,0,0,0,115,
        1,0,0,0,0,117,1,0,0,0,0,119,1,0,0,0,0,121,1,0,0,0,0,123,1,0,0,0,
        0,125,1,0,0,0,0,127,1,0,0,0,0,129,1,0,0,0,0,131,1,0,0,0,0,133,1,
        0,0,0,0,135,1,0,0,0,0,137,1,0,0,0,0,139,1,0,0,0,0,141,1,0,0,0,0,
        143,1,0,0,0,0,145,1,0,0,0,0,147,1,0,0,0,0,149,1,0,0,0,0,153,1,0,
        0,0,0,155,1,0,0,0,0,157,1,0,0,0,0,159,1,0,0,0,0,161,1,0,0,0,0,163,
        1,0,0,0,0,165,1,0,0,0,0,167,1,0,0,0,0,169,1,0,0,0,0,171,1,0,0,0,
        0,173,1,0,0,0,0,175,1,0,0,0,0,177,1,0,0,0,0,179,1,0,0,0,0,181,1,
        0,0,0,0,183,1,0,0,0,1,185,1,0,0,0,3,187,1,0,0,0,5,190,1,0,0,0,7,
        208,1,0,0,0,9,223,1,0,0,0,11,240,1,0,0,0,13,244,1,0,0,0,15,252,1,
        0,0,0,17,276,1,0,0,0,19,280,1,0,0,0,21,295,1,0,0,0,23,312,1,0,0,
        0,25,320,1,0,0,0,27,380,1,0,0,0,29,384,1,0,0,0,31,386,1,0,0,0,33,
        388,1,0,0,0,35,390,1,0,0,0,37,392,1,0,0,0,39,394,1,0,0,0,41,396,
        1,0,0,0,43,398,1,0,0,0,45,400,1,0,0,0,47,403,1,0,0,0,49,406,1,0,
        0,0,51,408,1,0,0,0,53,410,1,0,0,0,55,412,1,0,0,0,57,420,1,0,0,0,
        59,427,1,0,0,0,61,435,1,0,0,0,63,443,1,0,0,0,65,503,1,0,0,0,67,520,
        1,0,0,0,69,522,1,0,0,0,71,527,1,0,0,0,73,533,1,0,0,0,75,538,1,0,
        0,0,77,543,1,0,0,0,79,547,1,0,0,0,81,551,1,0,0,0,83,556,1,0,0,0,
        85,561,1,0,0,0,87,566,1,0,0,0,89,571,1,0,0,0,91,576,1,0,0,0,93,581,
        1,0,0,0,95,589,1,0,0,0,97,597,1,0,0,0,99,605,1,0,0,0,101,613,1,0,
        0,0,103,621,1,0,0,0,105,629,1,0,0,0,107,635,1,0,0,0,109,641,1,0,
        0,0,111,647,1,0,0,0,113,655,1,0,0,0,115,663,1,0,0,0,117,671,1,0,
        0,0,119,679,1,0,0,0,121,687,1,0,0,0,123,694,1,0,0,0,125,701,1,0,
        0,0,127,707,1,0,0,0,129,717,1,0,0,0,131,724,1,0,0,0,133,730,1,0,
        0,0,135,752,1,0,0,0,137,754,1,0,0,0,139,761,1,0,0,0,141,769,1,0,
        0,0,143,777,1,0,0,0,145,785,1,0,0,0,147,787,1,0,0,0,149,789,1,0,
        0,0,151,791,1,0,0,0,153,793,1,0,0,0,155,809,1,0,0,0,157,811,1,0,
        0,0,159,833,1,0,0,0,161,835,1,0,0,0,163,840,1,0,0,0,165,851,1,0,
        0,0,167,853,1,0,0,0,169,859,1,0,0,0,171,869,1,0,0,0,173,880,1,0,
        0,0,175,882,1,0,0,0,177,888,1,0,0,0,179,898,1,0,0,0,181,901,1,0,
        0,0,183,905,1,0,0,0,185,186,5,44,0,0,186,2,1,0,0,0,187,188,5,46,
        0,0,188,4,1,0,0,0,189,191,7,0,0,0,190,189,1,0,0,0,191,192,1,0,0,
        0,192,190,1,0,0,0,192,193,1,0,0,0,193,194,1,0,0,0,194,195,6,2,0,
        0,195,6,1,0,0,0,196,197,5,92,0,0,197,209,5,44,0,0,198,199,5,92,0,
        0,199,200,5,116,0,0,200,201,5,104,0,0,201,202,5,105,0,0,202,203,
        5,110,0,0,203,204,5,115,0,0,204,205,5,112,0,0,205,206,5,97,0,0,206,
        207,5,99,0,0,207,209,5,101,0,0,208,196,1,0,0,0,208,198,1,0,0,0,209,
        210,1,0,0,0,210,211,6,3,0,0,211,8,1,0,0,0,212,213,5,92,0,0,213,224,
        5,58,0,0,214,215,5,92,0,0,215,216,5,109,0,0,216,217,5,101,0,0,217,
        218,5,100,0,0,218,219,5,115,0,0,219,220,5,112,0,0,220,221,5,97,0,
        0,221,222,5,99,0,0,222,224,5,101,0,0,223,212,1,0,0,0,223,214,1,0,
        0,0,224,225,1,0,0,0,225,226,6,4,0,0,226,10,1,0,0,0,227,228,5,92,
        0,0,228,241,5,59,0,0,229,230,5,92,0,0,230,231,5,116,0,0,231,232,
        5,104,0,0,232,233,5,105,0,0,233,234,5,99,0,0,234,235,5,107,0,0,235,
        236,5,115,0,0,236,237,5,112,0,0,237,238,5,97,0,0,238,239,5,99,0,
        0,239,241,5,101,0,0,240,227,1,0,0,0,240,229,1,0,0,0,241,242,1,0,
        0,0,242,243,6,5,0,0,243,12,1,0,0,0,244,245,5,92,0,0,245,246,5,113,
        0,0,246,247,5,117,0,0,247,248,5,97,0,0,248,249,5,100,0,0,249,250,
        1,0,0,0,250,251,6,6,0,0,251,14,1,0,0,0,252,253,5,92,0,0,253,254,
        5,113,0,0,254,255,5,113,0,0,255,256,5,117,0,0,256,257,5,97,0,0,257,
        258,5,100,0,0,258,259,1,0,0,0,259,260,6,7,0,0,260,16,1,0,0,0,261,
        262,5,92,0,0,262,277,5,33,0,0,263,264,5,92,0,0,264,265,5,110,0,0,
        265,266,5,101,0,0,266,267,5,103,0,0,267,268,5,116,0,0,268,269,5,
        104,0,0,269,270,5,105,0,0,270,271,5,110,0,0,271,272,5,115,0,0,272,
        273,5,112,0,0,273,274,5,97,0,0,274,275,5,99,0,0,275,277,5,101,0,
        0,276,261,1,0,0,0,276,263,1,0,0,0,277,278,1,0,0,0,278,279,6,8,0,
        0,279,18,1,0,0,0,280,281,5,92,0,0,281,282,5,110,0,0,282,283,5,101,
        0,0,283,284,5,103,0,0,284,285,5,109,0,0,285,286,5,101,0,0,286,287,
        5,100,0,0,287,288,5,115,0,0,288,289,5,112,0,0,289,290,5,97,0,0,290,
        291,5,99,0,0,291,292,5,101,0,0,292,293,1,0,0,0,293,294,6,9,0,0,294,
        20,1,0,0,0,295,296,5,92,0,0,296,297,5,110,0,0,297,298,5,101,0,0,
        298,299,5,103,0,0,299,300,5,116,0,0,300,301,5,104,0,0,301,302,5,
        105,0,0,302,303,5,99,0,0,303,304,5,107,0,0,304,305,5,115,0,0,305,
        306,5,112,0,0,306,307,5,97,0,0,307,308,5,99,0,0,308,309,5,101,0,
        0,309,310,1,0,0,0,310,311,6,10,0,0,311,22,1,0,0,0,312,313,5,92,0,
        0,313,314,5,108,0,0,314,315,5,101,0,0,315,316,5,102,0,0,316,317,
        5,116,0,0,317,318,1,0,0,0,318,319,6,11,0,0,319,24,1,0,0,0,320,321,
        5,92,0,0,321,322,5,114,0,0,322,323,5,105,0,0,323,324,5,103,0,0,324,
        325,5,104,0,0,325,326,5,116,0,0,326,327,1,0,0,0,327,328,6,12,0,0,
        328,26,1,0,0,0,329,330,5,92,0,0,330,331,5,118,0,0,331,332,5,114,
        0,0,332,333,5,117,0,0,333,334,5,108,0,0,334,381,5,101,0,0,335,336,
        5,92,0,0,336,337,5,118,0,0,337,338,5,99,0,0,338,339,5,101,0,0,339,
        340,5,110,0,0,340,341,5,116,0,0,341,342,5,101,0,0,342,381,5,114,
        0,0,343,344,5,92,0,0,344,345,5,118,0,0,345,346,5,98,0,0,346,347,
        5,111,0,0,347,381,5,120,0,0,348,349,5,92,0,0,349,350,5,118,0,0,350,
        351,5,115,0,0,351,352,5,107,0,0,352,353,5,105,0,0,353,381,5,112,
        0,0,354,355,5,92,0,0,355,356,5,118,0,0,356,357,5,115,0,0,357,358,
        5,112,0,0,358,359,5,97,0,0,359,360,5,99,0,0,360,381,5,101,0,0,361,
        362,5,92,0,0,362,363,5,104,0,0,363,364,5,102,0,0,364,365,5,105,0,
        0,365,381,5,108,0,0,366,367,5,92,0,0,367,381,5,42,0,0,368,369,5,
        92,0,0,369,381,5,45,0,0,370,371,5,92,0,0,371,381,5,46,0,0,372,373,
        5,92,0,0,373,381,5,47,0,0,374,375,5,92,0,0,375,381,5,34,0,0,376,
        377,5,92,0,0,377,381,5,40,0,0,378,379,5,92,0,0,379,381,5,61,0,0,
        380,329,1,0,0,0,380,335,1,0,0,0,380,343,1,0,0,0,380,348,1,0,0,0,
        380,354,1,0,0,0,380,361,1,0,0,0,380,366,1,0,0,0,380,368,1,0,0,0,
        380,370,1,0,0,0,380,372,1,0,0,0,380,374,1,0,0,0,380,376,1,0,0,0,
        380,378,1,0,0,0,381,382,1,0,0,0,382,383,6,13,0,0,383,28,1,0,0,0,
        384,385,5,43,0,0,385,30,1,0,0,0,386,387,5,45,0,0,387,32,1,0,0,0,
        388,389,5,42,0,0,389,34,1,0,0,0,390,391,5,47,0,0,391,36,1,0,0,0,
        392,393,5,40,0,0,393,38,1,0,0,0,394,395,5,41,0,0,395,40,1,0,0,0,
        396,397,5,123,0,0,397,42,1,0,0,0,398,399,5,125,0,0,399,44,1,0,0,
        0,400,401,5,92,0,0,401,402,5,123,0,0,402,46,1,0,0,0,403,404,5,92,
        0,0,404,405,5,125,0,0,405,48,1,0,0,0,406,407,5,91,0,0,407,50,1,0,
        0,0,408,409,5,93,0,0,409,52,1,0,0,0,410,411,5,124,0,0,411,54,1,0,
        0,0,412,413,5,92,0,0,413,414,5,114,0,0,414,415,5,105,0,0,415,416,
        5,103,0,0,416,417,5,104,0,0,417,418,5,116,0,0,418,419,5,124,0,0,
        419,56,1,0,0,0,420,421,5,92,0,0,421,422,5,108,0,0,422,423,5,101,
        0,0,423,424,5,102,0,0,424,425,5,116,0,0,425,426,5,124,0,0,426,58,
        1,0,0,0,427,428,5,92,0,0,428,429,5,108,0,0,429,430,5,97,0,0,430,
        431,5,110,0,0,431,432,5,103,0,0,432,433,5,108,0,0,433,434,5,101,
        0,0,434,60,1,0,0,0,435,436,5,92,0,0,436,437,5,114,0,0,437,438,5,
        97,0,0,438,439,5,110,0,0,439,440,5,103,0,0,440,441,5,108,0,0,441,
        442,5,101,0,0,442,62,1,0,0,0,443,444,5,92,0,0,444,445,5,108,0,0,
        445,446,5,105,0,0,446,447,5,109,0,0,447,64,1,0,0,0,448,449,5,92,
        0,0,449,450,5,116,0,0,450,504,5,111,0,0,451,452,5,92,0,0,452,453,
        5,114,0,0,453,454,5,105,0,0,454,455,5,103,0,0,455,456,5,104,0,0,
        456,457,5,116,0,0,457,458,5,97,0,0,458,459,5,114,0,0,459,460,5,114,
        0,0,460,461,5,111,0,0,461,504,5,119,0,0,462,463,5,92,0,0,463,464,
        5,82,0,0,464,465,5,105,0,0,465,466,5,103,0,0,466,467,5,104,0,0,467,
        468,5,116,0,0,468,469,5,97,0,0,469,470,5,114,0,0,470,471,5,114,0,
        0,471,472,5,111,0,0,472,504,5,119,0,0,473,474,5,92,0,0,474,475,5,
        108,0,0,475,476,5,111,0,0,476,477,5,110,0,0,477,478,5,103,0,0,478,
        479,5,114,0,0,479,480,5,105,0,0,480,481,5,103,0,0,481,482,5,104,
        0,0,482,483,5,116,0,0,483,484,5,97,0,0,484,485,5,114,0,0,485,486,
        5,114,0,0,486,487,5,111,0,0,487,504,5,119,0,0,488,489,5,92,0,0,489,
        490,5,76,0,0,490,491,5,111,0,0,491,492,5,110,0,0,492,493,5,103,0,
        0,493,494,5,114,0,0,494,495,5,105,0,0,495,496,5,103,0,0,496,497,
        5,104,0,0,497,498,5,116,0,0,498,499,5,97,0,0,499,500,5,114,0,0,500,
        501,5,114,0,0,501,502,5,111,0,0,502,504,5,119,0,0,503,448,1,0,0,
        0,503,451,1,0,0,0,503,462,1,0,0,0,503,473,1,0,0,0,503,488,1,0,0,
        0,504,66,1,0,0,0,505,506,5,92,0,0,506,507,5,105,0,0,507,508,5,110,
        0,0,508,521,5,116,0,0,509,510,5,92,0,0,510,511,5,105,0,0,511,512,
        5,110,0,0,512,513,5,116,0,0,513,514,5,92,0,0,514,515,5,108,0,0,515,
        516,5,105,0,0,516,517,5,109,0,0,517,518,5,105,0,0,518,519,5,116,
        0,0,519,521,5,115,0,0,520,505,1,0,0,0,520,509,1,0,0,0,521,68,1,0,
        0,0,522,523,5,92,0,0,523,524,5,115,0,0,524,525,5,117,0,0,525,526,
        5,109,0,0,526,70,1,0,0,0,527,528,5,92,0,0,528,529,5,112,0,0,529,
        530,5,114,0,0,530,531,5,111,0,0,531,532,5,100,0,0,532,72,1,0,0,0,
        533,534,5,92,0,0,534,535,5,101,0,0,535,536,5,120,0,0,536,537,5,112,
        0,0,537,74,1,0,0,0,538,539,5,92,0,0,539,540,5,108,0,0,540,541,5,
        111,0,0,541,542,5,103,0,0,542,76,1,0,0,0,543,544,5,92,0,0,544,545,
        5,108,0,0,545,546,5,103,0,0,546,78,1,0,0,0,547,548,5,92,0,0,548,
        549,5,108,0,0,549,550,5,110,0,0,550,80,1,0,0,0,551,552,5,92,0,0,
        552,553,5,115,0,0,553,554,5,105,0,0,554,555,5,110,0,0,555,82,1,0,
        0,0,556,557,5,92,0,0,557,558,5,99,0,0,558,559,5,111,0,0,559,560,
        5,115,0,0,560,84,1,0,0,0,561,562,5,92,0,0,562,563,5,116,0,0,563,
        564,5,97,0,0,564,565,5,110,0,0,565,86,1,0,0,0,566,567,5,92,0,0,567,
        568,5,99,0,0,568,569,5,115,0,0,569,570,5,99,0,0,570,88,1,0,0,0,571,
        572,5,92,0,0,572,573,5,115,0,0,573,574,5,101,0,0,574,575,5,99,0,
        0,575,90,1,0,0,0,576,577,5,92,0,0,577,578,5,99,0,0,578,579,5,111,
        0,0,579,580,5,116,0,0,580,92,1,0,0,0,581,582,5,92,0,0,582,583,5,
        97,0,0,583,584,5,114,0,0,584,585,5,99,0,0,585,586,5,115,0,0,586,
        587,5,105,0,0,587,588,5,110,0,0,588,94,1,0,0,0,589,590,5,92,0,0,
        590,591,5,97,0,0,591,592,5,114,0,0,592,593,5,99,0,0,593,594,5,99,
        0,0,594,595,5,111,0,0,595,596,5,115,0,0,596,96,1,0,0,0,597,598,5,
        92,0,0,598,599,5,97,0,0,599,600,5,114,0,0,600,601,5,99,0,0,601,602,
        5,116,0,0,602,603,5,97,0,0,603,604,5,110,0,0,604,98,1,0,0,0,605,
        606,5,92,0,0,606,607,5,97,0,0,607,608,5,114,0,0,608,609,5,99,0,0,
        609,610,5,99,0,0,610,611,5,115,0,0,611,612,5,99,0,0,612,100,1,0,
        0,0,613,614,5,92,0,0,614,615,5,97,0,0,615,616,5,114,0,0,616,617,
        5,99,0,0,617,618,5,115,0,0,618,619,5,101,0,0,619,620,5,99,0,0,620,
        102,1,0,0,0,621,622,5,92,0,0,622,623,5,97,0,0,623,624,5,114,0,0,
        624,625,5,99,0,0,625,626,5,99,0,0,626,627,5,111,0,0,627,628,5,116,
        0,0,628,104,1,0,0,0,629,630,5,92,0,0,630,631,5,115,0,0,631,632,5,
        105,0,0,632,633,5,110,0,0,633,634,5,104,0,0,634,106,1,0,0,0,635,
        636,5,92,0,0,636,637,5,99,0,0,637,638,5,111,0,0,638,639,5,115,0,
        0,639,640,5,104,0,0,640,108,1,0,0,0,641,642,5,92,0,0,642,643,5,116,
        0,0,643,644,5,97,0,0,644,645,5,110,0,0,645,646,5,104,0,0,646,110,
        1,0,0,0,647,648,5,92,0,0,648,649,5,97,0,0,649,650,5,114,0,0,650,
        651,5,115,0,0,651,652,5,105,0,0,652,653,5,110,0,0,653,654,5,104,
        0,0,654,112,1,0,0,0,655,656,5,92,0,0,656,657,5,97,0,0,657,658,5,
        114,0,0,658,659,5,99,0,0,659,660,5,111,0,0,660,661,5,115,0,0,661,
        662,5,104,0,0,662,114,1,0,0,0,663,664,5,92,0,0,664,665,5,97,0,0,
        665,666,5,114,0,0,666,667,5,116,0,0,667,668,5,97,0,0,668,669,5,110,
        0,0,669,670,5,104,0,0,670,116,1,0,0,0,671,672,5,92,0,0,672,673,5,
        108,0,0,673,674,5,102,0,0,674,675,5,108,0,0,675,676,5,111,0,0,676,
        677,5,111,0,0,677,678,5,114,0,0,678,118,1,0,0,0,679,680,5,92,0,0,
        680,681,5,114,0,0,681,682,5,102,0,0,682,683,5,108,0,0,683,684,5,
        111,0,0,684,685,5,111,0,0,685,686,5,114,0,0,686,120,1,0,0,0,687,
        688,5,92,0,0,688,689,5,108,0,0,689,690,5,99,0,0,690,691,5,101,0,
        0,691,692,5,105,0,0,692,693,5,108,0,0,693,122,1,0,0,0,694,695,5,
        92,0,0,695,696,5,114,0,0,696,697,5,99,0,0,697,698,5,101,0,0,698,
        699,5,105,0,0,699,700,5,108,0,0,700,124,1,0,0,0,701,702,5,92,0,0,
        702,703,5,115,0,0,703,704,5,113,0,0,704,705,5,114,0,0,705,706,5,
        116,0,0,706,126,1,0,0,0,707,708,5,92,0,0,708,709,5,111,0,0,709,710,
        5,118,0,0,710,711,5,101,0,0,711,712,5,114,0,0,712,713,5,108,0,0,
        713,714,5,105,0,0,714,715,5,110,0,0,715,716,5,101,0,0,716,128,1,
        0,0,0,717,718,5,92,0,0,718,719,5,116,0,0,719,720,5,105,0,0,720,721,
        5,109,0,0,721,722,5,101,0,0,722,723,5,115,0,0,723,130,1,0,0,0,724,
        725,5,92,0,0,725,726,5,99,0,0,726,727,5,100,0,0,727,728,5,111,0,
        0,728,729,5,116,0,0,729,132,1,0,0,0,730,731,5,92,0,0,731,732,5,100,
        0,0,732,733,5,105,0,0,733,734,5,118,0,0,734,134,1,0,0,0,735,736,
        5,92,0,0,736,737,5,102,0,0,737,738,5,114,0,0,738,739,5,97,0,0,739,
        753,5,99,0,0,740,741,5,92,0,0,741,742,5,100,0,0,742,743,5,102,0,
        0,743,744,5,114,0,0,744,745,5,97,0,0,745,753,5,99,0,0,746,747,5,
        92,0,0,747,748,5,116,0,0,748,749,5,102,0,0,749,750,5,114,0,0,750,
        751,5,97,0,0,751,753,5,99,0,0,752,735,1,0,0,0,752,740,1,0,0,0,752,
        746,1,0,0,0,753,136,1,0,0,0,754,755,5,92,0,0,755,756,5,98,0,0,756,
        757,5,105,0,0,757,758,5,110,0,0,758,759,5,111,0,0,759,760,5,109,
        0,0,760,138,1,0,0,0,761,762,5,92,0,0,762,763,5,100,0,0,763,764,5,
        98,0,0,764,765,5,105,0,0,765,766,5,110,0,0,766,767,5,111,0,0,767,
        768,5,109,0,0,768,140,1,0,0,0,769,770,5,92,0,0,770,771,5,116,0,0,
        771,772,5,98,0,0,772,773,5,105,0,0,773,774,5,110,0,0,774,775,5,111,
        0,0,775,776,5,109,0,0,776,142,1,0,0,0,777,778,5,92,0,0,778,779,5,
        109,0,0,779,780,5,97,0,0,780,781,5,116,0,0,781,782,5,104,0,0,782,
        783,5,105,0,0,783,784,5,116,0,0,784,144,1,0,0,0,785,786,5,95,0,0,
        786,146,1,0,0,0,787,788,5,94,0,0,788,148,1,0,0,0,789,790,5,58,0,
        0,790,150,1,0,0,0,791,792,7,0,0,0,792,152,1,0,0,0,793,797,5,100,
        0,0,794,796,3,151,75,0,795,794,1,0,0,0,796,799,1,0,0,0,797,798,1,
        0,0,0,797,795,1,0,0,0,798,807,1,0,0,0,799,797,1,0,0,0,800,808,7,
        1,0,0,801,803,5,92,0,0,802,804,7,1,0,0,803,802,1,0,0,0,804,805,1,
        0,0,0,805,803,1,0,0,0,805,806,1,0,0,0,806,808,1,0,0,0,807,800,1,
        0,0,0,807,801,1,0,0,0,808,154,1,0,0,0,809,810,7,1,0,0,810,156,1,
        0,0,0,811,812,7,2,0,0,812,158,1,0,0,0,813,817,5,38,0,0,814,816,3,
        151,75,0,815,814,1,0,0,0,816,819,1,0,0,0,817,818,1,0,0,0,817,815,
        1,0,0,0,818,821,1,0,0,0,819,817,1,0,0,0,820,813,1,0,0,0,820,821,
        1,0,0,0,821,822,1,0,0,0,822,834,5,61,0,0,823,831,5,61,0,0,824,826,
        3,151,75,0,825,824,1,0,0,0,826,829,1,0,0,0,827,828,1,0,0,0,827,825,
        1,0,0,0,828,830,1,0,0,0,829,827,1,0,0,0,830,832,5,38,0,0,831,827,
        1,0,0,0,831,832,1,0,0,0,832,834,1,0,0,0,833,820,1,0,0,0,833,823,
        1,0,0,0,834,160,1,0,0,0,835,836,5,92,0,0,836,837,5,110,0,0,837,838,
        5,101,0,0,838,839,5,113,0,0,839,162,1,0,0,0,840,841,5,60,0,0,841,
        164,1,0,0,0,842,843,5,92,0,0,843,844,5,108,0,0,844,845,5,101,0,0,
        845,852,5,113,0,0,846,847,5,92,0,0,847,848,5,108,0,0,848,852,5,101,
        0,0,849,852,3,167,83,0,850,852,3,169,84,0,851,842,1,0,0,0,851,846,
        1,0,0,0,851,849,1,0,0,0,851,850,1,0,0,0,852,166,1,0,0,0,853,854,
        5,92,0,0,854,855,5,108,0,0,855,856,5,101,0,0,856,857,5,113,0,0,857,
        858,5,113,0,0,858,168,1,0,0,0,859,860,5,92,0,0,860,861,5,108,0,0,
        861,862,5,101,0,0,862,863,5,113,0,0,863,864,5,115,0,0,864,865,5,
        108,0,0,865,866,5,97,0,0,866,867,5,110,0,0,867,868,5,116,0,0,868,
        170,1,0,0,0,869,870,5,62,0,0,870,172,1,0,0,0,871,872,5,92,0,0,872,
        873,5,103,0,0,873,874,5,101,0,0,874,881,5,113,0,0,875,876,5,92,0,
        0,876,877,5,103,0,0,877,881,5,101,0,0,878,881,3,175,87,0,879,881,
        3,177,88,0,880,871,1,0,0,0,880,875,1,0,0,0,880,878,1,0,0,0,880,879,
        1,0,0,0,881,174,1,0,0,0,882,883,5,92,0,0,883,884,5,103,0,0,884,885,
        5,101,0,0,885,886,5,113,0,0,886,887,5,113,0,0,887,176,1,0,0,0,888,
        889,5,92,0,0,889,890,5,103,0,0,890,891,5,101,0,0,891,892,5,113,0,
        0,892,893,5,115,0,0,893,894,5,108,0,0,894,895,5,97,0,0,895,896,5,
        110,0,0,896,897,5,116,0,0,897,178,1,0,0,0,898,899,5,33,0,0,899,180,
        1,0,0,0,900,902,5,39,0,0,901,900,1,0,0,0,902,903,1,0,0,0,903,901,
        1,0,0,0,903,904,1,0,0,0,904,182,1,0,0,0,905,907,5,92,0,0,906,908,
        7,1,0,0,907,906,1,0,0,0,908,909,1,0,0,0,909,907,1,0,0,0,909,910,
        1,0,0,0,910,184,1,0,0,0,22,0,192,208,223,240,276,380,503,520,752,
        797,805,807,817,820,827,831,833,851,880,903,909,1,6,0,0
    ]

class LaTeXLexer(Lexer):

    atn = ATNDeserializer().deserialize(serializedATN())

    decisionsToDFA = [ DFA(ds, i) for i, ds in enumerate(atn.decisionToState) ]

    T__0 = 1
    T__1 = 2
    WS = 3
    THINSPACE = 4
    MEDSPACE = 5
    THICKSPACE = 6
    QUAD = 7
    QQUAD = 8
    NEGTHINSPACE = 9
    NEGMEDSPACE = 10
    NEGTHICKSPACE = 11
    CMD_LEFT = 12
    CMD_RIGHT = 13
    IGNORE = 14
    ADD = 15
    SUB = 16
    MUL = 17
    DIV = 18
    L_PAREN = 19
    R_PAREN = 20
    L_BRACE = 21
    R_BRACE = 22
    L_BRACE_LITERAL = 23
    R_BRACE_LITERAL = 24
    L_BRACKET = 25
    R_BRACKET = 26
    BAR = 27
    R_BAR = 28
    L_BAR = 29
    L_ANGLE = 30
    R_ANGLE = 31
    FUNC_LIM = 32
    LIM_APPROACH_SYM = 33
    FUNC_INT = 34
    FUNC_SUM = 35
    FUNC_PROD = 36
    FUNC_EXP = 37
    FUNC_LOG = 38
    FUNC_LG = 39
    FUNC_LN = 40
    FUNC_SIN = 41
    FUNC_COS = 42
    FUNC_TAN = 43
    FUNC_CSC = 44
    FUNC_SEC = 45
    FUNC_COT = 46
    FUNC_ARCSIN = 47
    FUNC_ARCCOS = 48
    FUNC_ARCTAN = 49
    FUNC_ARCCSC = 50
    FUNC_ARCSEC = 51
    FUNC_ARCCOT = 52
    FUNC_SINH = 53
    FUNC_COSH = 54
    FUNC_TANH = 55
    FUNC_ARSINH = 56
    FUNC_ARCOSH = 57
    FUNC_ARTANH = 58
    L_FLOOR = 59
    R_FLOOR = 60
    L_CEIL = 61
    R_CEIL = 62
    FUNC_SQRT = 63
    FUNC_OVERLINE = 64
    CMD_TIMES = 65
    CMD_CDOT = 66
    CMD_DIV = 67
    CMD_FRAC = 68
    CMD_BINOM = 69
    CMD_DBINOM = 70
    CMD_TBINOM = 71
    CMD_MATHIT = 72
    UNDERSCORE = 73
    CARET = 74
    COLON = 75
    DIFFERENTIAL = 76
    LETTER = 77
    DIGIT = 78
    EQUAL = 79
    NEQ = 80
    LT = 81
    LTE = 82
    LTE_Q = 83
    LTE_S = 84
    GT = 85
    GTE = 86
    GTE_Q = 87
    GTE_S = 88
    BANG = 89
    SINGLE_QUOTES = 90
    SYMBOL = 91

    channelNames = [ u"DEFAULT_TOKEN_CHANNEL", u"HIDDEN" ]

    modeNames = [ "DEFAULT_MODE" ]

    literalNames = [ "<INVALID>",
            "','", "'.'", "'\\quad'", "'\\qquad'", "'\\negmedspace'", "'\\negthickspace'",
            "'\\left'", "'\\right'", "'+'", "'-'", "'*'", "'/'", "'('",
            "')'", "'{'", "'}'", "'\\{'", "'\\}'", "'['", "']'", "'|'",
            "'\\right|'", "'\\left|'", "'\\langle'", "'\\rangle'", "'\\lim'",
            "'\\sum'", "'\\prod'", "'\\exp'", "'\\log'", "'\\lg'", "'\\ln'",
            "'\\sin'", "'\\cos'", "'\\tan'", "'\\csc'", "'\\sec'", "'\\cot'",
            "'\\arcsin'", "'\\arccos'", "'\\arctan'", "'\\arccsc'", "'\\arcsec'",
            "'\\arccot'", "'\\sinh'", "'\\cosh'", "'\\tanh'", "'\\arsinh'",
            "'\\arcosh'", "'\\artanh'", "'\\lfloor'", "'\\rfloor'", "'\\lceil'",
            "'\\rceil'", "'\\sqrt'", "'\\overline'", "'\\times'", "'\\cdot'",
            "'\\div'", "'\\binom'", "'\\dbinom'", "'\\tbinom'", "'\\mathit'",
            "'_'", "'^'", "':'", "'\\neq'", "'<'", "'\\leqq'", "'\\leqslant'",
            "'>'", "'\\geqq'", "'\\geqslant'", "'!'" ]

    symbolicNames = [ "<INVALID>",
            "WS", "THINSPACE", "MEDSPACE", "THICKSPACE", "QUAD", "QQUAD",
            "NEGTHINSPACE", "NEGMEDSPACE", "NEGTHICKSPACE", "CMD_LEFT",
            "CMD_RIGHT", "IGNORE", "ADD", "SUB", "MUL", "DIV", "L_PAREN",
            "R_PAREN", "L_BRACE", "R_BRACE", "L_BRACE_LITERAL", "R_BRACE_LITERAL",
            "L_BRACKET", "R_BRACKET", "BAR", "R_BAR", "L_BAR", "L_ANGLE",
            "R_ANGLE", "FUNC_LIM", "LIM_APPROACH_SYM", "FUNC_INT", "FUNC_SUM",
            "FUNC_PROD", "FUNC_EXP", "FUNC_LOG", "FUNC_LG", "FUNC_LN", "FUNC_SIN",
            "FUNC_COS", "FUNC_TAN", "FUNC_CSC", "FUNC_SEC", "FUNC_COT",
            "FUNC_ARCSIN", "FUNC_ARCCOS", "FUNC_ARCTAN", "FUNC_ARCCSC",
            "FUNC_ARCSEC", "FUNC_ARCCOT", "FUNC_SINH", "FUNC_COSH", "FUNC_TANH",
            "FUNC_ARSINH", "FUNC_ARCOSH", "FUNC_ARTANH", "L_FLOOR", "R_FLOOR",
            "L_CEIL", "R_CEIL", "FUNC_SQRT", "FUNC_OVERLINE", "CMD_TIMES",
            "CMD_CDOT", "CMD_DIV", "CMD_FRAC", "CMD_BINOM", "CMD_DBINOM",
            "CMD_TBINOM", "CMD_MATHIT", "UNDERSCORE", "CARET", "COLON",
            "DIFFERENTIAL", "LETTER", "DIGIT", "EQUAL", "NEQ", "LT", "LTE",
            "LTE_Q", "LTE_S", "GT", "GTE", "GTE_Q", "GTE_S", "BANG", "SINGLE_QUOTES",
            "SYMBOL" ]

    ruleNames = [ "T__0", "T__1", "WS", "THINSPACE", "MEDSPACE", "THICKSPACE",
                  "QUAD", "QQUAD", "NEGTHINSPACE", "NEGMEDSPACE", "NEGTHICKSPACE",
                  "CMD_LEFT", "CMD_RIGHT", "IGNORE", "ADD", "SUB", "MUL",
                  "DIV", "L_PAREN", "R_PAREN", "L_BRACE", "R_BRACE", "L_BRACE_LITERAL",
                  "R_BRACE_LITERAL", "L_BRACKET", "R_BRACKET", "BAR", "R_BAR",
                  "L_BAR", "L_ANGLE", "R_ANGLE", "FUNC_LIM", "LIM_APPROACH_SYM",
                  "FUNC_INT", "FUNC_SUM", "FUNC_PROD", "FUNC_EXP", "FUNC_LOG",
                  "FUNC_LG", "FUNC_LN", "FUNC_SIN", "FUNC_COS", "FUNC_TAN",
                  "FUNC_CSC", "FUNC_SEC", "FUNC_COT", "FUNC_ARCSIN", "FUNC_ARCCOS",
                  "FUNC_ARCTAN", "FUNC_ARCCSC", "FUNC_ARCSEC", "FUNC_ARCCOT",
                  "FUNC_SINH", "FUNC_COSH", "FUNC_TANH", "FUNC_ARSINH",
                  "FUNC_ARCOSH", "FUNC_ARTANH", "L_FLOOR", "R_FLOOR", "L_CEIL",
                  "R_CEIL", "FUNC_SQRT", "FUNC_OVERLINE", "CMD_TIMES", "CMD_CDOT",
                  "CMD_DIV", "CMD_FRAC", "CMD_BINOM", "CMD_DBINOM", "CMD_TBINOM",
                  "CMD_MATHIT", "UNDERSCORE", "CARET", "COLON", "WS_CHAR",
                  "DIFFERENTIAL", "LETTER", "DIGIT", "EQUAL", "NEQ", "LT",
                  "LTE", "LTE_Q", "LTE_S", "GT", "GTE", "GTE_Q", "GTE_S",
                  "BANG", "SINGLE_QUOTES", "SYMBOL" ]

    grammarFileName = "LaTeX.g4"

    def __init__(self, input=None, output:TextIO = sys.stdout):
        super().__init__(input, output)
        self.checkVersion("4.11.1")
        self._interp = LexerATNSimulator(self, self.atn, self.decisionsToDFA, PredictionContextCache())
        self._actions = None
        self._predicates = None



# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\mechanics\lagrange.py ===
from sympy import diff, zeros, Matrix, eye, sympify
from sympy.core.sorting import default_sort_key
from sympy.physics.vector import dynamicsymbols, ReferenceFrame
from sympy.physics.mechanics.method import _Methods
from sympy.physics.mechanics.functions import (
    find_dynamicsymbols, msubs, _f_list_parser, _validate_coordinates)
from sympy.physics.mechanics.linearize import Linearizer
from sympy.utilities.iterables import iterable

__all__ = ['LagrangesMethod']


class LagrangesMethod(_Methods):
    """Lagrange's method object.

    Explanation
    ===========

    This object generates the equations of motion in a two step procedure. The
    first step involves the initialization of LagrangesMethod by supplying the
    Lagrangian and the generalized coordinates, at the bare minimum. If there
    are any constraint equations, they can be supplied as keyword arguments.
    The Lagrange multipliers are automatically generated and are equal in
    number to the constraint equations. Similarly any non-conservative forces
    can be supplied in an iterable (as described below and also shown in the
    example) along with a ReferenceFrame. This is also discussed further in the
    __init__ method.

    Attributes
    ==========

    q, u : Matrix
        Matrices of the generalized coordinates and speeds
    loads : iterable
        Iterable of (Point, vector) or (ReferenceFrame, vector) tuples
        describing the forces on the system.
    bodies : iterable
        Iterable containing the rigid bodies and particles of the system.
    mass_matrix : Matrix
        The system's mass matrix
    forcing : Matrix
        The system's forcing vector
    mass_matrix_full : Matrix
        The "mass matrix" for the qdot's, qdoubledot's, and the
        lagrange multipliers (lam)
    forcing_full : Matrix
        The forcing vector for the qdot's, qdoubledot's and
        lagrange multipliers (lam)

    Examples
    ========

    This is a simple example for a one degree of freedom translational
    spring-mass-damper.

    In this example, we first need to do the kinematics.
    This involves creating generalized coordinates and their derivatives.
    Then we create a point and set its velocity in a frame.

        >>> from sympy.physics.mechanics import LagrangesMethod, Lagrangian
        >>> from sympy.physics.mechanics import ReferenceFrame, Particle, Point
        >>> from sympy.physics.mechanics import dynamicsymbols
        >>> from sympy import symbols
        >>> q = dynamicsymbols('q')
        >>> qd = dynamicsymbols('q', 1)
        >>> m, k, b = symbols('m k b')
        >>> N = ReferenceFrame('N')
        >>> P = Point('P')
        >>> P.set_vel(N, qd * N.x)

    We need to then prepare the information as required by LagrangesMethod to
    generate equations of motion.
    First we create the Particle, which has a point attached to it.
    Following this the lagrangian is created from the kinetic and potential
    energies.
    Then, an iterable of nonconservative forces/torques must be constructed,
    where each item is a (Point, Vector) or (ReferenceFrame, Vector) tuple,
    with the Vectors representing the nonconservative forces or torques.

        >>> Pa = Particle('Pa', P, m)
        >>> Pa.potential_energy = k * q**2 / 2.0
        >>> L = Lagrangian(N, Pa)
        >>> fl = [(P, -b * qd * N.x)]

    Finally we can generate the equations of motion.
    First we create the LagrangesMethod object. To do this one must supply
    the Lagrangian, and the generalized coordinates. The constraint equations,
    the forcelist, and the inertial frame may also be provided, if relevant.
    Next we generate Lagrange's equations of motion, such that:
    Lagrange's equations of motion = 0.
    We have the equations of motion at this point.

        >>> l = LagrangesMethod(L, [q], forcelist = fl, frame = N)
        >>> print(l.form_lagranges_equations())
        Matrix([[b*Derivative(q(t), t) + 1.0*k*q(t) + m*Derivative(q(t), (t, 2))]])

    We can also solve for the states using the 'rhs' method.

        >>> print(l.rhs())
        Matrix([[Derivative(q(t), t)], [(-b*Derivative(q(t), t) - 1.0*k*q(t))/m]])

    Please refer to the docstrings on each method for more details.
    """

    def __init__(self, Lagrangian, qs, forcelist=None, bodies=None, frame=None,
                 hol_coneqs=None, nonhol_coneqs=None):
        """Supply the following for the initialization of LagrangesMethod.

        Lagrangian : Sympifyable

        qs : array_like
            The generalized coordinates

        hol_coneqs : array_like, optional
            The holonomic constraint equations

        nonhol_coneqs : array_like, optional
            The nonholonomic constraint equations

        forcelist : iterable, optional
            Takes an iterable of (Point, Vector) or (ReferenceFrame, Vector)
            tuples which represent the force at a point or torque on a frame.
            This feature is primarily to account for the nonconservative forces
            and/or moments.

        bodies : iterable, optional
            Takes an iterable containing the rigid bodies and particles of the
            system.

        frame : ReferenceFrame, optional
            Supply the inertial frame. This is used to determine the
            generalized forces due to non-conservative forces.
        """

        self._L = Matrix([sympify(Lagrangian)])
        self.eom = None
        self._m_cd = Matrix()           # Mass Matrix of differentiated coneqs
        self._m_d = Matrix()            # Mass Matrix of dynamic equations
        self._f_cd = Matrix()           # Forcing part of the diff coneqs
        self._f_d = Matrix()            # Forcing part of the dynamic equations
        self.lam_coeffs = Matrix()      # The coeffecients of the multipliers

        forcelist = forcelist if forcelist else []
        if not iterable(forcelist):
            raise TypeError('Force pairs must be supplied in an iterable.')
        self._forcelist = forcelist
        if frame and not isinstance(frame, ReferenceFrame):
            raise TypeError('frame must be a valid ReferenceFrame')
        self._bodies = bodies
        self.inertial = frame

        self.lam_vec = Matrix()

        self._term1 = Matrix()
        self._term2 = Matrix()
        self._term3 = Matrix()
        self._term4 = Matrix()

        # Creating the qs, qdots and qdoubledots
        if not iterable(qs):
            raise TypeError('Generalized coordinates must be an iterable')
        self._q = Matrix(qs)
        self._qdots = self.q.diff(dynamicsymbols._t)
        self._qdoubledots = self._qdots.diff(dynamicsymbols._t)
        _validate_coordinates(self.q)

        mat_build = lambda x: Matrix(x) if x else Matrix()
        hol_coneqs = mat_build(hol_coneqs)
        nonhol_coneqs = mat_build(nonhol_coneqs)
        self.coneqs = Matrix([hol_coneqs.diff(dynamicsymbols._t),
                nonhol_coneqs])
        self._hol_coneqs = hol_coneqs

    def form_lagranges_equations(self):
        """Method to form Lagrange's equations of motion.

        Returns a vector of equations of motion using Lagrange's equations of
        the second kind.
        """

        qds = self._qdots
        qdd_zero = dict.fromkeys(self._qdoubledots, 0)
        n = len(self.q)

        # Internally we represent the EOM as four terms:
        # EOM = term1 - term2 - term3 - term4 = 0

        # First term
        self._term1 = self._L.jacobian(qds)
        self._term1 = self._term1.diff(dynamicsymbols._t).T

        # Second term
        self._term2 = self._L.jacobian(self.q).T

        # Third term
        if self.coneqs:
            coneqs = self.coneqs
            m = len(coneqs)
            # Creating the multipliers
            self.lam_vec = Matrix(dynamicsymbols('lam1:' + str(m + 1)))
            self.lam_coeffs = -coneqs.jacobian(qds)
            self._term3 = self.lam_coeffs.T * self.lam_vec
            # Extracting the coeffecients of the qdds from the diff coneqs
            diffconeqs = coneqs.diff(dynamicsymbols._t)
            self._m_cd = diffconeqs.jacobian(self._qdoubledots)
            # The remaining terms i.e. the 'forcing' terms in diff coneqs
            self._f_cd = -diffconeqs.subs(qdd_zero)
        else:
            self._term3 = zeros(n, 1)

        # Fourth term
        if self.forcelist:
            N = self.inertial
            self._term4 = zeros(n, 1)
            for i, qd in enumerate(qds):
                flist = zip(*_f_list_parser(self.forcelist, N))
                self._term4[i] = sum(v.diff(qd, N).dot(f) for (v, f) in flist)
        else:
            self._term4 = zeros(n, 1)

        # Form the dynamic mass and forcing matrices
        without_lam = self._term1 - self._term2 - self._term4
        self._m_d = without_lam.jacobian(self._qdoubledots)
        self._f_d = -without_lam.subs(qdd_zero)

        # Form the EOM
        self.eom = without_lam - self._term3
        return self.eom

    def _form_eoms(self):
        return self.form_lagranges_equations()

    @property
    def mass_matrix(self):
        """Returns the mass matrix, which is augmented by the Lagrange
        multipliers, if necessary.

        Explanation
        ===========

        If the system is described by 'n' generalized coordinates and there are
        no constraint equations then an n X n matrix is returned.

        If there are 'n' generalized coordinates and 'm' constraint equations
        have been supplied during initialization then an n X (n+m) matrix is
        returned. The (n + m - 1)th and (n + m)th columns contain the
        coefficients of the Lagrange multipliers.
        """

        if self.eom is None:
            raise ValueError('Need to compute the equations of motion first')
        if self.coneqs:
            return (self._m_d).row_join(self.lam_coeffs.T)
        else:
            return self._m_d

    @property
    def mass_matrix_full(self):
        """Augments the coefficients of qdots to the mass_matrix."""

        if self.eom is None:
            raise ValueError('Need to compute the equations of motion first')
        n = len(self.q)
        m = len(self.coneqs)
        row1 = eye(n).row_join(zeros(n, n + m))
        row2 = zeros(n, n).row_join(self.mass_matrix)
        if self.coneqs:
            row3 = zeros(m, n).row_join(self._m_cd).row_join(zeros(m, m))
            return row1.col_join(row2).col_join(row3)
        else:
            return row1.col_join(row2)

    @property
    def forcing(self):
        """Returns the forcing vector from 'lagranges_equations' method."""

        if self.eom is None:
            raise ValueError('Need to compute the equations of motion first')
        return self._f_d

    @property
    def forcing_full(self):
        """Augments qdots to the forcing vector above."""

        if self.eom is None:
            raise ValueError('Need to compute the equations of motion first')
        if self.coneqs:
            return self._qdots.col_join(self.forcing).col_join(self._f_cd)
        else:
            return self._qdots.col_join(self.forcing)

    def to_linearizer(self, q_ind=None, qd_ind=None, q_dep=None, qd_dep=None,
                      linear_solver='LU'):
        """Returns an instance of the Linearizer class, initiated from the data
        in the LagrangesMethod class. This may be more desirable than using the
        linearize class method, as the Linearizer object will allow more
        efficient recalculation (i.e. about varying operating points).

        Parameters
        ==========

        q_ind, qd_ind : array_like, optional
            The independent generalized coordinates and speeds.
        q_dep, qd_dep : array_like, optional
            The dependent generalized coordinates and speeds.
        linear_solver : str, callable
            Method used to solve the several symbolic linear systems of the
            form ``A*x=b`` in the linearization process. If a string is
            supplied, it should be a valid method that can be used with the
            :meth:`sympy.matrices.matrixbase.MatrixBase.solve`. If a callable is
            supplied, it should have the format ``x = f(A, b)``, where it
            solves the equations and returns the solution. The default is
            ``'LU'`` which corresponds to SymPy's ``A.LUsolve(b)``.
            ``LUsolve()`` is fast to compute but will often result in
            divide-by-zero and thus ``nan`` results.

        Returns
        =======
        Linearizer
            An instantiated
            :class:`sympy.physics.mechanics.linearize.Linearizer`.

        """

        # Compose vectors
        t = dynamicsymbols._t
        q = self.q
        u = self._qdots
        ud = u.diff(t)
        # Get vector of lagrange multipliers
        lams = self.lam_vec

        mat_build = lambda x: Matrix(x) if x else Matrix()
        q_i = mat_build(q_ind)
        q_d = mat_build(q_dep)
        u_i = mat_build(qd_ind)
        u_d = mat_build(qd_dep)

        # Compose general form equations
        f_c = self._hol_coneqs
        f_v = self.coneqs
        f_a = f_v.diff(t)
        f_0 = u
        f_1 = -u
        f_2 = self._term1
        f_3 = -(self._term2 + self._term4)
        f_4 = -self._term3

        # Check that there are an appropriate number of independent and
        # dependent coordinates
        if len(q_d) != len(f_c) or len(u_d) != len(f_v):
            raise ValueError(("Must supply {:} dependent coordinates, and " +
                    "{:} dependent speeds").format(len(f_c), len(f_v)))
        if set(Matrix([q_i, q_d])) != set(q):
            raise ValueError("Must partition q into q_ind and q_dep, with " +
                    "no extra or missing symbols.")
        if set(Matrix([u_i, u_d])) != set(u):
            raise ValueError("Must partition qd into qd_ind and qd_dep, " +
                    "with no extra or missing symbols.")

        # Find all other dynamic symbols, forming the forcing vector r.
        # Sort r to make it canonical.
        insyms = set(Matrix([q, u, ud, lams]))
        r = list(find_dynamicsymbols(f_3, insyms))
        r.sort(key=default_sort_key)
        # Check for any derivatives of variables in r that are also found in r.
        for i in r:
            if diff(i, dynamicsymbols._t) in r:
                raise ValueError('Cannot have derivatives of specified \
                                 quantities when linearizing forcing terms.')

        return Linearizer(f_0, f_1, f_2, f_3, f_4, f_c, f_v, f_a, q, u, q_i,
                          q_d, u_i, u_d, r, lams, linear_solver=linear_solver)

    def linearize(self, q_ind=None, qd_ind=None, q_dep=None, qd_dep=None,
                  linear_solver='LU', **kwargs):
        """Linearize the equations of motion about a symbolic operating point.

        Parameters
        ==========
        linear_solver : str, callable
            Method used to solve the several symbolic linear systems of the
            form ``A*x=b`` in the linearization process. If a string is
            supplied, it should be a valid method that can be used with the
            :meth:`sympy.matrices.matrixbase.MatrixBase.solve`. If a callable is
            supplied, it should have the format ``x = f(A, b)``, where it
            solves the equations and returns the solution. The default is
            ``'LU'`` which corresponds to SymPy's ``A.LUsolve(b)``.
            ``LUsolve()`` is fast to compute but will often result in
            divide-by-zero and thus ``nan`` results.
        **kwargs
            Extra keyword arguments are passed to
            :meth:`sympy.physics.mechanics.linearize.Linearizer.linearize`.

        Explanation
        ===========

        If kwarg A_and_B is False (default), returns M, A, B, r for the
        linearized form, M*[q', u']^T = A*[q_ind, u_ind]^T + B*r.

        If kwarg A_and_B is True, returns A, B, r for the linearized form
        dx = A*x + B*r, where x = [q_ind, u_ind]^T. Note that this is
        computationally intensive if there are many symbolic parameters. For
        this reason, it may be more desirable to use the default A_and_B=False,
        returning M, A, and B. Values may then be substituted in to these
        matrices, and the state space form found as
        A = P.T*M.inv()*A, B = P.T*M.inv()*B, where P = Linearizer.perm_mat.

        In both cases, r is found as all dynamicsymbols in the equations of
        motion that are not part of q, u, q', or u'. They are sorted in
        canonical form.

        The operating points may be also entered using the ``op_point`` kwarg.
        This takes a dictionary of {symbol: value}, or a an iterable of such
        dictionaries. The values may be numeric or symbolic. The more values
        you can specify beforehand, the faster this computation will run.

        For more documentation, please see the ``Linearizer`` class."""

        linearizer = self.to_linearizer(q_ind, qd_ind, q_dep, qd_dep,
                                        linear_solver=linear_solver)
        result = linearizer.linearize(**kwargs)
        return result + (linearizer.r,)

    def solve_multipliers(self, op_point=None, sol_type='dict'):
        """Solves for the values of the lagrange multipliers symbolically at
        the specified operating point.

        Parameters
        ==========

        op_point : dict or iterable of dicts, optional
            Point at which to solve at. The operating point is specified as
            a dictionary or iterable of dictionaries of {symbol: value}. The
            value may be numeric or symbolic itself.

        sol_type : str, optional
            Solution return type. Valid options are:
            - 'dict': A dict of {symbol : value} (default)
            - 'Matrix': An ordered column matrix of the solution
        """

        # Determine number of multipliers
        k = len(self.lam_vec)
        if k == 0:
            raise ValueError("System has no lagrange multipliers to solve for.")
        # Compose dict of operating conditions
        if isinstance(op_point, dict):
            op_point_dict = op_point
        elif iterable(op_point):
            op_point_dict = {}
            for op in op_point:
                op_point_dict.update(op)
        elif op_point is None:
            op_point_dict = {}
        else:
            raise TypeError("op_point must be either a dictionary or an "
                            "iterable of dictionaries.")
        # Compose the system to be solved
        mass_matrix = self.mass_matrix.col_join(-self.lam_coeffs.row_join(
                zeros(k, k)))
        force_matrix = self.forcing.col_join(self._f_cd)
        # Sub in the operating point
        mass_matrix = msubs(mass_matrix, op_point_dict)
        force_matrix = msubs(force_matrix, op_point_dict)
        # Solve for the multipliers
        sol_list = mass_matrix.LUsolve(-force_matrix)[-k:]
        if sol_type == 'dict':
            return dict(zip(self.lam_vec, sol_list))
        elif sol_type == 'Matrix':
            return Matrix(sol_list)
        else:
            raise ValueError("Unknown sol_type {:}.".format(sol_type))

    def rhs(self, inv_method=None, **kwargs):
        """Returns equations that can be solved numerically.

        Parameters
        ==========

        inv_method : str
            The specific sympy inverse matrix calculation method to use. For a
            list of valid methods, see
            :meth:`~sympy.matrices.matrixbase.MatrixBase.inv`
        """

        if inv_method is None:
            self._rhs = self.mass_matrix_full.LUsolve(self.forcing_full)
        else:
            self._rhs = (self.mass_matrix_full.inv(inv_method,
                         try_block_diag=True) * self.forcing_full)
        return self._rhs

    @property
    def q(self):
        return self._q

    @property
    def u(self):
        return self._qdots

    @property
    def bodies(self):
        return self._bodies

    @property
    def forcelist(self):
        return self._forcelist

    @property
    def loads(self):
        return self._forcelist

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\stats\frv.py ===
"""
Finite Discrete Random Variables Module

See Also
========
sympy.stats.frv_types
sympy.stats.rv
sympy.stats.crv
"""
from itertools import product

from sympy.concrete.summations import Sum
from sympy.core.basic import Basic
from sympy.core.cache import cacheit
from sympy.core.function import Lambda
from sympy.core.mul import Mul
from sympy.core.numbers import (I, nan)
from sympy.core.relational import Eq
from sympy.core.singleton import S
from sympy.core.symbol import (Dummy, Symbol)
from sympy.core.sympify import sympify
from sympy.functions.elementary.exponential import exp
from sympy.functions.elementary.piecewise import Piecewise
from sympy.logic.boolalg import (And, Or)
from sympy.sets.sets import Intersection
from sympy.core.containers import Dict
from sympy.core.logic import Logic
from sympy.core.relational import Relational
from sympy.core.sympify import _sympify
from sympy.sets.sets import FiniteSet
from sympy.stats.rv import (RandomDomain, ProductDomain, ConditionalDomain,
                            PSpace, IndependentProductPSpace, SinglePSpace, random_symbols,
                            sumsets, rv_subs, NamedArgsMixin, Density, Distribution)


class FiniteDensity(dict):
    """
    A domain with Finite Density.
    """
    def __call__(self, item):
        """
        Make instance of a class callable.

        If item belongs to current instance of a class, return it.

        Otherwise, return 0.
        """
        item = sympify(item)
        if item in self:
            return self[item]
        else:
            return 0

    @property
    def dict(self):
        """
        Return item as dictionary.
        """
        return dict(self)

class FiniteDomain(RandomDomain):
    """
    A domain with discrete finite support

    Represented using a FiniteSet.
    """
    is_Finite = True

    @property
    def symbols(self):
        return FiniteSet(sym for sym, val in self.elements)

    @property
    def elements(self):
        return self.args[0]

    @property
    def dict(self):
        return FiniteSet(*[Dict(dict(el)) for el in self.elements])

    def __contains__(self, other):
        return other in self.elements

    def __iter__(self):
        return self.elements.__iter__()

    def as_boolean(self):
        return Or(*[And(*[Eq(sym, val) for sym, val in item]) for item in self])


class SingleFiniteDomain(FiniteDomain):
    """
    A FiniteDomain over a single symbol/set

    Example: The possibilities of a *single* die roll.
    """

    def __new__(cls, symbol, set):
        if not isinstance(set, FiniteSet) and \
            not isinstance(set, Intersection):
            set = FiniteSet(*set)
        return Basic.__new__(cls, symbol, set)

    @property
    def symbol(self):
        return self.args[0]

    @property
    def symbols(self):
        return FiniteSet(self.symbol)

    @property
    def set(self):
        return self.args[1]

    @property
    def elements(self):
        return FiniteSet(*[frozenset(((self.symbol, elem), )) for elem in self.set])

    def __iter__(self):
        return (frozenset(((self.symbol, elem),)) for elem in self.set)

    def __contains__(self, other):
        sym, val = tuple(other)[0]
        return sym == self.symbol and val in self.set


class ProductFiniteDomain(ProductDomain, FiniteDomain):
    """
    A Finite domain consisting of several other FiniteDomains

    Example: The possibilities of the rolls of three independent dice
    """

    def __iter__(self):
        proditer = product(*self.domains)
        return (sumsets(items) for items in proditer)

    @property
    def elements(self):
        return FiniteSet(*self)


class ConditionalFiniteDomain(ConditionalDomain, ProductFiniteDomain):
    """
    A FiniteDomain that has been restricted by a condition

    Example: The possibilities of a die roll under the condition that the
    roll is even.
    """

    def __new__(cls, domain, condition):
        """
        Create a new instance of ConditionalFiniteDomain class
        """
        if condition is True:
            return domain
        cond = rv_subs(condition)
        return Basic.__new__(cls, domain, cond)

    def _test(self, elem):
        """
        Test the value. If value is boolean, return it. If value is equality
        relational (two objects are equal), return it with left-hand side
        being equal to right-hand side. Otherwise, raise ValueError exception.
        """
        val = self.condition.xreplace(dict(elem))
        if val in [True, False]:
            return val
        elif val.is_Equality:
            return val.lhs == val.rhs
        raise ValueError("Undecidable if %s" % str(val))

    def __contains__(self, other):
        return other in self.fulldomain and self._test(other)

    def __iter__(self):
        return (elem for elem in self.fulldomain if self._test(elem))

    @property
    def set(self):
        if isinstance(self.fulldomain, SingleFiniteDomain):
            return FiniteSet(*[elem for elem in self.fulldomain.set
                               if frozenset(((self.fulldomain.symbol, elem),)) in self])
        else:
            raise NotImplementedError(
                "Not implemented on multi-dimensional conditional domain")

    def as_boolean(self):
        return FiniteDomain.as_boolean(self)


class SingleFiniteDistribution(Distribution, NamedArgsMixin):
    def __new__(cls, *args):
        args = list(map(sympify, args))
        return Basic.__new__(cls, *args)

    @staticmethod
    def check(*args):
        pass

    @property # type: ignore
    @cacheit
    def dict(self):
        if self.is_symbolic:
            return Density(self)
        return {k: self.pmf(k) for k in self.set}

    def pmf(self, *args): # to be overridden by specific distribution
        raise NotImplementedError()

    @property
    def set(self): # to be overridden by specific distribution
        raise NotImplementedError()

    values = property(lambda self: self.dict.values)
    items = property(lambda self: self.dict.items)
    is_symbolic = property(lambda self: False)
    __iter__ = property(lambda self: self.dict.__iter__)
    __getitem__ = property(lambda self: self.dict.__getitem__)

    def __call__(self, *args):
        return self.pmf(*args)

    def __contains__(self, other):
        return other in self.set


#=============================================
#=========  Probability Space  ===============
#=============================================


class FinitePSpace(PSpace):
    """
    A Finite Probability Space

    Represents the probabilities of a finite number of events.
    """
    is_Finite = True

    def __new__(cls, domain, density):
        density = {sympify(key): sympify(val)
                for key, val in density.items()}
        public_density = Dict(density)

        obj = PSpace.__new__(cls, domain, public_density)
        obj._density = density
        return obj

    def prob_of(self, elem):
        elem = sympify(elem)
        density = self._density
        if isinstance(list(density.keys())[0], FiniteSet):
            return density.get(elem, S.Zero)
        return density.get(tuple(elem)[0][1], S.Zero)

    def where(self, condition):
        assert all(r.symbol in self.symbols for r in random_symbols(condition))
        return ConditionalFiniteDomain(self.domain, condition)

    def compute_density(self, expr):
        expr = rv_subs(expr, self.values)
        d = FiniteDensity()
        for elem in self.domain:
            val = expr.xreplace(dict(elem))
            prob = self.prob_of(elem)
            d[val] = d.get(val, S.Zero) + prob
        return d

    @cacheit
    def compute_cdf(self, expr):
        d = self.compute_density(expr)
        cum_prob = S.Zero
        cdf = []
        for key in sorted(d):
            prob = d[key]
            cum_prob += prob
            cdf.append((key, cum_prob))

        return dict(cdf)

    @cacheit
    def sorted_cdf(self, expr, python_float=False):
        cdf = self.compute_cdf(expr)
        items = list(cdf.items())
        sorted_items = sorted(items, key=lambda val_cumprob: val_cumprob[1])
        if python_float:
            sorted_items = [(v, float(cum_prob))
                    for v, cum_prob in sorted_items]
        return sorted_items

    @cacheit
    def compute_characteristic_function(self, expr):
        d = self.compute_density(expr)
        t = Dummy('t', real=True)

        return Lambda(t, sum(exp(I*k*t)*v for k,v in d.items()))

    @cacheit
    def compute_moment_generating_function(self, expr):
        d = self.compute_density(expr)
        t = Dummy('t', real=True)

        return Lambda(t, sum(exp(k*t)*v for k,v in d.items()))

    def compute_expectation(self, expr, rvs=None, **kwargs):
        rvs = rvs or self.values
        expr = rv_subs(expr, rvs)
        probs = [self.prob_of(elem) for elem in self.domain]
        if isinstance(expr, (Logic, Relational)):
            parse_domain = [tuple(elem)[0][1] for elem in self.domain]
            bools = [expr.xreplace(dict(elem)) for elem in self.domain]
        else:
            parse_domain = [expr.xreplace(dict(elem)) for elem in self.domain]
            bools = [True for elem in self.domain]
        return sum(Piecewise((prob * elem, blv), (S.Zero, True))
                for prob, elem, blv in zip(probs, parse_domain, bools))

    def compute_quantile(self, expr):
        cdf = self.compute_cdf(expr)
        p = Dummy('p', real=True)
        set = ((nan, (p < 0) | (p > 1)),)
        for key, value in cdf.items():
            set = set + ((key, p <= value), )
        return Lambda(p, Piecewise(*set))

    def probability(self, condition):
        cond_symbols = frozenset(rs.symbol for rs in random_symbols(condition))
        cond = rv_subs(condition)
        if not cond_symbols.issubset(self.symbols):
            raise ValueError("Cannot compare foreign random symbols, %s"
                             %(str(cond_symbols - self.symbols)))
        if isinstance(condition, Relational) and \
            (not cond.free_symbols.issubset(self.domain.free_symbols)):
            rv = condition.lhs if isinstance(condition.rhs, Symbol) else condition.rhs
            return sum(Piecewise(
                       (self.prob_of(elem), condition.subs(rv, list(elem)[0][1])),
                       (S.Zero, True)) for elem in self.domain)
        return sympify(sum(self.prob_of(elem) for elem in self.where(condition)))

    def conditional_space(self, condition):
        domain = self.where(condition)
        prob = self.probability(condition)
        density = {key: val / prob
                for key, val in self._density.items() if domain._test(key)}
        return FinitePSpace(domain, density)

    def sample(self, size=(), library='scipy', seed=None):
        """
        Internal sample method

        Returns dictionary mapping RandomSymbol to realization value.
        """
        return {self.value: self.distribution.sample(size, library, seed)}


class SingleFinitePSpace(SinglePSpace, FinitePSpace):
    """
    A single finite probability space

    Represents the probabilities of a set of random events that can be
    attributed to a single variable/symbol.

    This class is implemented by many of the standard FiniteRV types such as
    Die, Bernoulli, Coin, etc....
    """
    @property
    def domain(self):
        return SingleFiniteDomain(self.symbol, self.distribution.set)

    @property
    def _is_symbolic(self):
        """
        Helper property to check if the distribution
        of the random variable is having symbolic
        dimension.
        """
        return self.distribution.is_symbolic

    @property
    def distribution(self):
        return self.args[1]

    def pmf(self, expr):
        return self.distribution.pmf(expr)

    @property # type: ignore
    @cacheit
    def _density(self):
        return {FiniteSet((self.symbol, val)): prob
                    for val, prob in self.distribution.dict.items()}

    @cacheit
    def compute_characteristic_function(self, expr):
        if self._is_symbolic:
            d = self.compute_density(expr)
            t = Dummy('t', real=True)
            ki = Dummy('ki')
            return Lambda(t, Sum(d(ki)*exp(I*ki*t), (ki, self.args[1].low, self.args[1].high)))
        expr = rv_subs(expr, self.values)
        return FinitePSpace(self.domain, self.distribution).compute_characteristic_function(expr)

    @cacheit
    def compute_moment_generating_function(self, expr):
        if self._is_symbolic:
            d = self.compute_density(expr)
            t = Dummy('t', real=True)
            ki = Dummy('ki')
            return Lambda(t, Sum(d(ki)*exp(ki*t), (ki, self.args[1].low, self.args[1].high)))
        expr = rv_subs(expr, self.values)
        return FinitePSpace(self.domain, self.distribution).compute_moment_generating_function(expr)

    def compute_quantile(self, expr):
        if self._is_symbolic:
            raise NotImplementedError("Computing quantile for random variables "
            "with symbolic dimension because the bounds of searching the required "
            "value is undetermined.")
        expr = rv_subs(expr, self.values)
        return FinitePSpace(self.domain, self.distribution).compute_quantile(expr)

    def compute_density(self, expr):
        if self._is_symbolic:
            rv = list(random_symbols(expr))[0]
            k = Dummy('k', integer=True)
            cond = True if not isinstance(expr, (Relational, Logic)) \
                     else expr.subs(rv, k)
            return Lambda(k,
            Piecewise((self.pmf(k), And(k >= self.args[1].low,
            k <= self.args[1].high, cond)), (S.Zero, True)))
        expr = rv_subs(expr, self.values)
        return FinitePSpace(self.domain, self.distribution).compute_density(expr)

    def compute_cdf(self, expr):
        if self._is_symbolic:
            d = self.compute_density(expr)
            k = Dummy('k')
            ki = Dummy('ki')
            return Lambda(k, Sum(d(ki), (ki, self.args[1].low, k)))
        expr = rv_subs(expr, self.values)
        return FinitePSpace(self.domain, self.distribution).compute_cdf(expr)

    def compute_expectation(self, expr, rvs=None, **kwargs):
        if self._is_symbolic:
            rv = random_symbols(expr)[0]
            k = Dummy('k', integer=True)
            expr = expr.subs(rv, k)
            cond = True if not isinstance(expr, (Relational, Logic)) \
                    else expr
            func = self.pmf(k) * k if cond != True else self.pmf(k) * expr
            return Sum(Piecewise((func, cond), (S.Zero, True)),
                (k, self.distribution.low, self.distribution.high)).doit()

        expr = _sympify(expr)
        expr = rv_subs(expr, rvs)
        return FinitePSpace(self.domain, self.distribution).compute_expectation(expr, rvs, **kwargs)

    def probability(self, condition):
        if self._is_symbolic:
            #TODO: Implement the mechanism for handling queries for symbolic sized distributions.
            raise NotImplementedError("Currently, probability queries are not "
            "supported for random variables with symbolic sized distributions.")
        condition = rv_subs(condition)
        return FinitePSpace(self.domain, self.distribution).probability(condition)

    def conditional_space(self, condition):
        """
        This method is used for transferring the
        computation to probability method because
        conditional space of random variables with
        symbolic dimensions is currently not possible.
        """
        if self._is_symbolic:
            self
        domain = self.where(condition)
        prob = self.probability(condition)
        density = {key: val / prob
                for key, val in self._density.items() if domain._test(key)}
        return FinitePSpace(domain, density)


class ProductFinitePSpace(IndependentProductPSpace, FinitePSpace):
    """
    A collection of several independent finite probability spaces
    """
    @property
    def domain(self):
        return ProductFiniteDomain(*[space.domain for space in self.spaces])

    @property  # type: ignore
    @cacheit
    def _density(self):
        proditer = product(*[iter(space._density.items())
            for space in self.spaces])
        d = {}
        for items in proditer:
            elems, probs = list(zip(*items))
            elem = sumsets(elems)
            prob = Mul(*probs)
            d[elem] = d.get(elem, S.Zero) + prob
        return Dict(d)

    @property  # type: ignore
    @cacheit
    def density(self):
        return Dict(self._density)

    def probability(self, condition):
        return FinitePSpace.probability(self, condition)

    def compute_density(self, expr):
        return FinitePSpace.compute_density(self, expr)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\cachecontrol\controller.py ===
# SPDX-FileCopyrightText: 2015 Eric Larson
#
# SPDX-License-Identifier: Apache-2.0

"""
The httplib2 algorithms ported for use with requests.
"""

from __future__ import annotations

import calendar
import logging
import re
import time
import weakref
from email.utils import parsedate_tz
from typing import TYPE_CHECKING, Collection, Mapping

from pip._vendor.requests.structures import CaseInsensitiveDict

from pip._vendor.cachecontrol.cache import DictCache, SeparateBodyBaseCache
from pip._vendor.cachecontrol.serialize import Serializer

if TYPE_CHECKING:
    from typing import Literal

    from pip._vendor.requests import PreparedRequest
    from pip._vendor.urllib3 import HTTPResponse

    from pip._vendor.cachecontrol.cache import BaseCache

logger = logging.getLogger(__name__)

URI = re.compile(r"^(([^:/?#]+):)?(//([^/?#]*))?([^?#]*)(\?([^#]*))?(#(.*))?")

PERMANENT_REDIRECT_STATUSES = (301, 308)


def parse_uri(uri: str) -> tuple[str, str, str, str, str]:
    """Parses a URI using the regex given in Appendix B of RFC 3986.

    (scheme, authority, path, query, fragment) = parse_uri(uri)
    """
    match = URI.match(uri)
    assert match is not None
    groups = match.groups()
    return (groups[1], groups[3], groups[4], groups[6], groups[8])


class CacheController:
    """An interface to see if request should cached or not."""

    def __init__(
        self,
        cache: BaseCache | None = None,
        cache_etags: bool = True,
        serializer: Serializer | None = None,
        status_codes: Collection[int] | None = None,
    ):
        self.cache = DictCache() if cache is None else cache
        self.cache_etags = cache_etags
        self.serializer = serializer or Serializer()
        self.cacheable_status_codes = status_codes or (200, 203, 300, 301, 308)

    @classmethod
    def _urlnorm(cls, uri: str) -> str:
        """Normalize the URL to create a safe key for the cache"""
        (scheme, authority, path, query, fragment) = parse_uri(uri)
        if not scheme or not authority:
            raise Exception("Only absolute URIs are allowed. uri = %s" % uri)

        scheme = scheme.lower()
        authority = authority.lower()

        if not path:
            path = "/"

        # Could do syntax based normalization of the URI before
        # computing the digest. See Section 6.2.2 of Std 66.
        request_uri = query and "?".join([path, query]) or path
        defrag_uri = scheme + "://" + authority + request_uri

        return defrag_uri

    @classmethod
    def cache_url(cls, uri: str) -> str:
        return cls._urlnorm(uri)

    def parse_cache_control(self, headers: Mapping[str, str]) -> dict[str, int | None]:
        known_directives = {
            # https://tools.ietf.org/html/rfc7234#section-5.2
            "max-age": (int, True),
            "max-stale": (int, False),
            "min-fresh": (int, True),
            "no-cache": (None, False),
            "no-store": (None, False),
            "no-transform": (None, False),
            "only-if-cached": (None, False),
            "must-revalidate": (None, False),
            "public": (None, False),
            "private": (None, False),
            "proxy-revalidate": (None, False),
            "s-maxage": (int, True),
        }

        cc_headers = headers.get("cache-control", headers.get("Cache-Control", ""))

        retval: dict[str, int | None] = {}

        for cc_directive in cc_headers.split(","):
            if not cc_directive.strip():
                continue

            parts = cc_directive.split("=", 1)
            directive = parts[0].strip()

            try:
                typ, required = known_directives[directive]
            except KeyError:
                logger.debug("Ignoring unknown cache-control directive: %s", directive)
                continue

            if not typ or not required:
                retval[directive] = None
            if typ:
                try:
                    retval[directive] = typ(parts[1].strip())
                except IndexError:
                    if required:
                        logger.debug(
                            "Missing value for cache-control " "directive: %s",
                            directive,
                        )
                except ValueError:
                    logger.debug(
                        "Invalid value for cache-control directive " "%s, must be %s",
                        directive,
                        typ.__name__,
                    )

        return retval

    def _load_from_cache(self, request: PreparedRequest) -> HTTPResponse | None:
        """
        Load a cached response, or return None if it's not available.
        """
        # We do not support caching of partial content: so if the request contains a
        # Range header then we don't want to load anything from the cache.
        if "Range" in request.headers:
            return None

        cache_url = request.url
        assert cache_url is not None
        cache_data = self.cache.get(cache_url)
        if cache_data is None:
            logger.debug("No cache entry available")
            return None

        if isinstance(self.cache, SeparateBodyBaseCache):
            body_file = self.cache.get_body(cache_url)
        else:
            body_file = None

        result = self.serializer.loads(request, cache_data, body_file)
        if result is None:
            logger.warning("Cache entry deserialization failed, entry ignored")
        return result

    def cached_request(self, request: PreparedRequest) -> HTTPResponse | Literal[False]:
        """
        Return a cached response if it exists in the cache, otherwise
        return False.
        """
        assert request.url is not None
        cache_url = self.cache_url(request.url)
        logger.debug('Looking up "%s" in the cache', cache_url)
        cc = self.parse_cache_control(request.headers)

        # Bail out if the request insists on fresh data
        if "no-cache" in cc:
            logger.debug('Request header has "no-cache", cache bypassed')
            return False

        if "max-age" in cc and cc["max-age"] == 0:
            logger.debug('Request header has "max_age" as 0, cache bypassed')
            return False

        # Check whether we can load the response from the cache:
        resp = self._load_from_cache(request)
        if not resp:
            return False

        # If we have a cached permanent redirect, return it immediately. We
        # don't need to test our response for other headers b/c it is
        # intrinsically "cacheable" as it is Permanent.
        #
        # See:
        #   https://tools.ietf.org/html/rfc7231#section-6.4.2
        #
        # Client can try to refresh the value by repeating the request
        # with cache busting headers as usual (ie no-cache).
        if int(resp.status) in PERMANENT_REDIRECT_STATUSES:
            msg = (
                "Returning cached permanent redirect response "
                "(ignoring date and etag information)"
            )
            logger.debug(msg)
            return resp

        headers: CaseInsensitiveDict[str] = CaseInsensitiveDict(resp.headers)
        if not headers or "date" not in headers:
            if "etag" not in headers:
                # Without date or etag, the cached response can never be used
                # and should be deleted.
                logger.debug("Purging cached response: no date or etag")
                self.cache.delete(cache_url)
            logger.debug("Ignoring cached response: no date")
            return False

        now = time.time()
        time_tuple = parsedate_tz(headers["date"])
        assert time_tuple is not None
        date = calendar.timegm(time_tuple[:6])
        current_age = max(0, now - date)
        logger.debug("Current age based on date: %i", current_age)

        # TODO: There is an assumption that the result will be a
        #       urllib3 response object. This may not be best since we
        #       could probably avoid instantiating or constructing the
        #       response until we know we need it.
        resp_cc = self.parse_cache_control(headers)

        # determine freshness
        freshness_lifetime = 0

        # Check the max-age pragma in the cache control header
        max_age = resp_cc.get("max-age")
        if max_age is not None:
            freshness_lifetime = max_age
            logger.debug("Freshness lifetime from max-age: %i", freshness_lifetime)

        # If there isn't a max-age, check for an expires header
        elif "expires" in headers:
            expires = parsedate_tz(headers["expires"])
            if expires is not None:
                expire_time = calendar.timegm(expires[:6]) - date
                freshness_lifetime = max(0, expire_time)
                logger.debug("Freshness lifetime from expires: %i", freshness_lifetime)

        # Determine if we are setting freshness limit in the
        # request. Note, this overrides what was in the response.
        max_age = cc.get("max-age")
        if max_age is not None:
            freshness_lifetime = max_age
            logger.debug(
                "Freshness lifetime from request max-age: %i", freshness_lifetime
            )

        min_fresh = cc.get("min-fresh")
        if min_fresh is not None:
            # adjust our current age by our min fresh
            current_age += min_fresh
            logger.debug("Adjusted current age from min-fresh: %i", current_age)

        # Return entry if it is fresh enough
        if freshness_lifetime > current_age:
            logger.debug('The response is "fresh", returning cached response')
            logger.debug("%i > %i", freshness_lifetime, current_age)
            return resp

        # we're not fresh. If we don't have an Etag, clear it out
        if "etag" not in headers:
            logger.debug('The cached response is "stale" with no etag, purging')
            self.cache.delete(cache_url)

        # return the original handler
        return False

    def conditional_headers(self, request: PreparedRequest) -> dict[str, str]:
        resp = self._load_from_cache(request)
        new_headers = {}

        if resp:
            headers: CaseInsensitiveDict[str] = CaseInsensitiveDict(resp.headers)

            if "etag" in headers:
                new_headers["If-None-Match"] = headers["ETag"]

            if "last-modified" in headers:
                new_headers["If-Modified-Since"] = headers["Last-Modified"]

        return new_headers

    def _cache_set(
        self,
        cache_url: str,
        request: PreparedRequest,
        response: HTTPResponse,
        body: bytes | None = None,
        expires_time: int | None = None,
    ) -> None:
        """
        Store the data in the cache.
        """
        if isinstance(self.cache, SeparateBodyBaseCache):
            # We pass in the body separately; just put a placeholder empty
            # string in the metadata.
            self.cache.set(
                cache_url,
                self.serializer.dumps(request, response, b""),
                expires=expires_time,
            )
            # body is None can happen when, for example, we're only updating
            # headers, as is the case in update_cached_response().
            if body is not None:
                self.cache.set_body(cache_url, body)
        else:
            self.cache.set(
                cache_url,
                self.serializer.dumps(request, response, body),
                expires=expires_time,
            )

    def cache_response(
        self,
        request: PreparedRequest,
        response_or_ref: HTTPResponse | weakref.ReferenceType[HTTPResponse],
        body: bytes | None = None,
        status_codes: Collection[int] | None = None,
    ) -> None:
        """
        Algorithm for caching requests.

        This assumes a requests Response object.
        """
        if isinstance(response_or_ref, weakref.ReferenceType):
            response = response_or_ref()
            if response is None:
                # The weakref can be None only in case the user used streamed request
                # and did not consume or close it, and holds no reference to requests.Response.
                # In such case, we don't want to cache the response.
                return
        else:
            response = response_or_ref

        # From httplib2: Don't cache 206's since we aren't going to
        #                handle byte range requests
        cacheable_status_codes = status_codes or self.cacheable_status_codes
        if response.status not in cacheable_status_codes:
            logger.debug(
                "Status code %s not in %s", response.status, cacheable_status_codes
            )
            return

        response_headers: CaseInsensitiveDict[str] = CaseInsensitiveDict(
            response.headers
        )

        if "date" in response_headers:
            time_tuple = parsedate_tz(response_headers["date"])
            assert time_tuple is not None
            date = calendar.timegm(time_tuple[:6])
        else:
            date = 0

        # If we've been given a body, our response has a Content-Length, that
        # Content-Length is valid then we can check to see if the body we've
        # been given matches the expected size, and if it doesn't we'll just
        # skip trying to cache it.
        if (
            body is not None
            and "content-length" in response_headers
            and response_headers["content-length"].isdigit()
            and int(response_headers["content-length"]) != len(body)
        ):
            return

        cc_req = self.parse_cache_control(request.headers)
        cc = self.parse_cache_control(response_headers)

        assert request.url is not None
        cache_url = self.cache_url(request.url)
        logger.debug('Updating cache with response from "%s"', cache_url)

        # Delete it from the cache if we happen to have it stored there
        no_store = False
        if "no-store" in cc:
            no_store = True
            logger.debug('Response header has "no-store"')
        if "no-store" in cc_req:
            no_store = True
            logger.debug('Request header has "no-store"')
        if no_store and self.cache.get(cache_url):
            logger.debug('Purging existing cache entry to honor "no-store"')
            self.cache.delete(cache_url)
        if no_store:
            return

        # https://tools.ietf.org/html/rfc7234#section-4.1:
        # A Vary header field-value of "*" always fails to match.
        # Storing such a response leads to a deserialization warning
        # during cache lookup and is not allowed to ever be served,
        # so storing it can be avoided.
        if "*" in response_headers.get("vary", ""):
            logger.debug('Response header has "Vary: *"')
            return

        # If we've been given an etag, then keep the response
        if self.cache_etags and "etag" in response_headers:
            expires_time = 0
            if response_headers.get("expires"):
                expires = parsedate_tz(response_headers["expires"])
                if expires is not None:
                    expires_time = calendar.timegm(expires[:6]) - date

            expires_time = max(expires_time, 14 * 86400)

            logger.debug(f"etag object cached for {expires_time} seconds")
            logger.debug("Caching due to etag")
            self._cache_set(cache_url, request, response, body, expires_time)

        # Add to the cache any permanent redirects. We do this before looking
        # that the Date headers.
        elif int(response.status) in PERMANENT_REDIRECT_STATUSES:
            logger.debug("Caching permanent redirect")
            self._cache_set(cache_url, request, response, b"")

        # Add to the cache if the response headers demand it. If there
        # is no date header then we can't do anything about expiring
        # the cache.
        elif "date" in response_headers:
            time_tuple = parsedate_tz(response_headers["date"])
            assert time_tuple is not None
            date = calendar.timegm(time_tuple[:6])
            # cache when there is a max-age > 0
            max_age = cc.get("max-age")
            if max_age is not None and max_age > 0:
                logger.debug("Caching b/c date exists and max-age > 0")
                expires_time = max_age
                self._cache_set(
                    cache_url,
                    request,
                    response,
                    body,
                    expires_time,
                )

            # If the request can expire, it means we should cache it
            # in the meantime.
            elif "expires" in response_headers:
                if response_headers["expires"]:
                    expires = parsedate_tz(response_headers["expires"])
                    if expires is not None:
                        expires_time = calendar.timegm(expires[:6]) - date
                    else:
                        expires_time = None

                    logger.debug(
                        "Caching b/c of expires header. expires in {} seconds".format(
                            expires_time
                        )
                    )
                    self._cache_set(
                        cache_url,
                        request,
                        response,
                        body,
                        expires_time,
                    )

    def update_cached_response(
        self, request: PreparedRequest, response: HTTPResponse
    ) -> HTTPResponse:
        """On a 304 we will get a new set of headers that we want to
        update our cached value with, assuming we have one.

        This should only ever be called when we've sent an ETag and
        gotten a 304 as the response.
        """
        assert request.url is not None
        cache_url = self.cache_url(request.url)
        cached_response = self._load_from_cache(request)

        if not cached_response:
            # we didn't have a cached response
            return response

        # Lets update our headers with the headers from the new request:
        # http://tools.ietf.org/html/draft-ietf-httpbis-p4-conditional-26#section-4.1
        #
        # The server isn't supposed to send headers that would make
        # the cached body invalid. But... just in case, we'll be sure
        # to strip out ones we know that might be problmatic due to
        # typical assumptions.
        excluded_headers = ["content-length"]

        cached_response.headers.update(
            {
                k: v
                for k, v in response.headers.items()
                if k.lower() not in excluded_headers
            }
        )

        # we want a 200 b/c we have content via the cache
        cached_response.status = 200

        # update our cache
        self._cache_set(cache_url, request, cached_response)

        return cached_response