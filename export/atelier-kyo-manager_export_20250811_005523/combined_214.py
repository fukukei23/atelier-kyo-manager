
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\stats\tests\test_rv.py ===
from sympy.concrete.summations import Sum
from sympy.core.basic import Basic
from sympy.core.containers import Tuple
from sympy.core.function import Lambda
from sympy.core.numbers import (Rational, nan, oo, pi)
from sympy.core.relational import Eq
from sympy.core.singleton import S
from sympy.core.symbol import (Symbol, symbols)
from sympy.functions.combinatorial.factorials import (FallingFactorial, binomial)
from sympy.functions.elementary.exponential import (exp, log)
from sympy.functions.elementary.trigonometric import (cos, sin)
from sympy.functions.special.delta_functions import DiracDelta
from sympy.integrals.integrals import integrate
from sympy.logic.boolalg import (And, Or)
from sympy.matrices.dense import Matrix
from sympy.sets.sets import Interval
from sympy.tensor.indexed import Indexed
from sympy.stats import (Die, Normal, Exponential, FiniteRV, P, E, H, variance,
        density, given, independent, dependent, where, pspace, GaussianUnitaryEnsemble,
        random_symbols, sample, Geometric, factorial_moment, Binomial, Hypergeometric,
        DiscreteUniform, Poisson, characteristic_function, moment_generating_function,
        BernoulliProcess, Variance, Expectation, Probability, Covariance, covariance, cmoment,
        moment, median)
from sympy.stats.rv import (IndependentProductPSpace, rs_swap, Density, NamedArgsMixin,
        RandomSymbol, sample_iter, PSpace, is_random, RandomIndexedSymbol, RandomMatrixSymbol)
from sympy.testing.pytest import raises, skip, XFAIL, warns_deprecated_sympy
from sympy.external import import_module
from sympy.core.numbers import comp
from sympy.stats.frv_types import BernoulliDistribution
from sympy.core.symbol import Dummy
from sympy.functions.elementary.piecewise import Piecewise

def test_where():
    X, Y = Die('X'), Die('Y')
    Z = Normal('Z', 0, 1)

    assert where(Z**2 <= 1).set == Interval(-1, 1)
    assert where(Z**2 <= 1).as_boolean() == Interval(-1, 1).as_relational(Z.symbol)
    assert where(And(X > Y, Y > 4)).as_boolean() == And(
        Eq(X.symbol, 6), Eq(Y.symbol, 5))

    assert len(where(X < 3).set) == 2
    assert 1 in where(X < 3).set

    X, Y = Normal('X', 0, 1), Normal('Y', 0, 1)
    assert where(And(X**2 <= 1, X >= 0)).set == Interval(0, 1)
    XX = given(X, And(X**2 <= 1, X >= 0))
    assert XX.pspace.domain.set == Interval(0, 1)
    assert XX.pspace.domain.as_boolean() == \
        And(0 <= X.symbol, X.symbol**2 <= 1, -oo < X.symbol, X.symbol < oo)

    with raises(TypeError):
        XX = given(X, X + 3)


def test_random_symbols():
    X, Y = Normal('X', 0, 1), Normal('Y', 0, 1)

    assert set(random_symbols(2*X + 1)) == {X}
    assert set(random_symbols(2*X + Y)) == {X, Y}
    assert set(random_symbols(2*X + Y.symbol)) == {X}
    assert set(random_symbols(2)) == set()


def test_characteristic_function():
    #  Imports I from sympy
    from sympy.core.numbers import I
    X = Normal('X',0,1)
    Y = DiscreteUniform('Y', [1,2,7])
    Z = Poisson('Z', 2)
    t = symbols('_t')
    P = Lambda(t, exp(-t**2/2))
    Q = Lambda(t, exp(7*t*I)/3 + exp(2*t*I)/3 + exp(t*I)/3)
    R = Lambda(t, exp(2 * exp(t*I) - 2))


    assert characteristic_function(X).dummy_eq(P)
    assert characteristic_function(Y).dummy_eq(Q)
    assert characteristic_function(Z).dummy_eq(R)


def test_moment_generating_function():

    X = Normal('X',0,1)
    Y = DiscreteUniform('Y', [1,2,7])
    Z = Poisson('Z', 2)
    t = symbols('_t')
    P = Lambda(t, exp(t**2/2))
    Q = Lambda(t, (exp(7*t)/3 + exp(2*t)/3 + exp(t)/3))
    R = Lambda(t, exp(2 * exp(t) - 2))


    assert moment_generating_function(X).dummy_eq(P)
    assert moment_generating_function(Y).dummy_eq(Q)
    assert moment_generating_function(Z).dummy_eq(R)

def test_sample_iter():

    X = Normal('X',0,1)
    Y = DiscreteUniform('Y', [1, 2, 7])
    Z = Poisson('Z', 2)

    scipy = import_module('scipy')
    if not scipy:
        skip('Scipy is not installed. Abort tests')
    expr = X**2 + 3
    iterator = sample_iter(expr)

    expr2 = Y**2 + 5*Y + 4
    iterator2 = sample_iter(expr2)

    expr3 = Z**3 + 4
    iterator3 = sample_iter(expr3)

    def is_iterator(obj):
        if (
            hasattr(obj, '__iter__') and
            (hasattr(obj, 'next') or
            hasattr(obj, '__next__')) and
            callable(obj.__iter__) and
            obj.__iter__() is obj
           ):
            return True
        else:
            return False
    assert is_iterator(iterator)
    assert is_iterator(iterator2)
    assert is_iterator(iterator3)

def test_pspace():
    X, Y = Normal('X', 0, 1), Normal('Y', 0, 1)
    x = Symbol('x')

    raises(ValueError, lambda: pspace(5 + 3))
    raises(ValueError, lambda: pspace(x < 1))
    assert pspace(X) == X.pspace
    assert pspace(2*X + 1) == X.pspace
    assert pspace(2*X + Y) == IndependentProductPSpace(Y.pspace, X.pspace)

def test_rs_swap():
    X = Normal('x', 0, 1)
    Y = Exponential('y', 1)

    XX = Normal('x', 0, 2)
    YY = Normal('y', 0, 3)

    expr = 2*X + Y
    assert expr.subs(rs_swap((X, Y), (YY, XX))) == 2*XX + YY


def test_RandomSymbol():

    X = Normal('x', 0, 1)
    Y = Normal('x', 0, 2)
    assert X.symbol == Y.symbol
    assert X != Y

    assert X.name == X.symbol.name

    X = Normal('lambda', 0, 1) # make sure we can use protected terms
    X = Normal('Lambda', 0, 1) # make sure we can use SymPy terms


def test_RandomSymbol_diff():
    X = Normal('x', 0, 1)
    assert (2*X).diff(X)


def test_random_symbol_no_pspace():
    x = RandomSymbol(Symbol('x'))
    assert x.pspace == PSpace()

def test_overlap():
    X = Normal('x', 0, 1)
    Y = Normal('x', 0, 2)

    raises(ValueError, lambda: P(X > Y))


def test_IndependentProductPSpace():
    X = Normal('X', 0, 1)
    Y = Normal('Y', 0, 1)
    px = X.pspace
    py = Y.pspace
    assert pspace(X + Y) == IndependentProductPSpace(px, py)
    assert pspace(X + Y) == IndependentProductPSpace(py, px)


def test_E():
    assert E(5) == 5


def test_H():
    X = Normal('X', 0, 1)
    D = Die('D', sides = 4)
    G = Geometric('G', 0.5)
    assert H(X, X > 0) == -log(2)/2 + S.Half + log(pi)/2
    assert H(D, D > 2) == log(2)
    assert comp(H(G).evalf().round(2), 1.39)


def test_Sample():
    X = Die('X', 6)
    Y = Normal('Y', 0, 1)
    z = Symbol('z', integer=True)

    scipy = import_module('scipy')
    if not scipy:
        skip('Scipy is not installed. Abort tests')
    assert sample(X) in [1, 2, 3, 4, 5, 6]
    assert isinstance(sample(X + Y), float)

    assert P(X + Y > 0, Y < 0, numsamples=10).is_number
    assert E(X + Y, numsamples=10).is_number
    assert E(X**2 + Y, numsamples=10).is_number
    assert E((X + Y)**2, numsamples=10).is_number
    assert variance(X + Y, numsamples=10).is_number

    raises(TypeError, lambda: P(Y > z, numsamples=5))

    assert P(sin(Y) <= 1, numsamples=10) == 1.0
    assert P(sin(Y) <= 1, cos(Y) < 1, numsamples=10) == 1.0

    assert all(i in range(1, 7) for i in density(X, numsamples=10))
    assert all(i in range(4, 7) for i in density(X, X>3, numsamples=10))

    numpy = import_module('numpy')
    if not numpy:
        skip('Numpy is not installed. Abort tests')
    #Test Issue #21563: Output of sample must be a float or array
    assert isinstance(sample(X), (numpy.int32, numpy.int64))
    assert isinstance(sample(Y), numpy.float64)
    assert isinstance(sample(X, size=2), numpy.ndarray)

    with warns_deprecated_sympy():
        sample(X, numsamples=2)

@XFAIL
def test_samplingE():
    scipy = import_module('scipy')
    if not scipy:
        skip('Scipy is not installed. Abort tests')
    Y = Normal('Y', 0, 1)
    z = Symbol('z', integer=True)
    assert E(Sum(1/z**Y, (z, 1, oo)), Y > 2, numsamples=3).is_number


def test_given():
    X = Normal('X', 0, 1)
    Y = Normal('Y', 0, 1)
    A = given(X, True)
    B = given(X, Y > 2)

    assert X == A == B


def test_factorial_moment():
    X = Poisson('X', 2)
    Y = Binomial('Y', 2, S.Half)
    Z = Hypergeometric('Z', 4, 2, 2)
    assert factorial_moment(X, 2) == 4
    assert factorial_moment(Y, 2) == S.Half
    assert factorial_moment(Z, 2) == Rational(1, 3)

    x, y, z, l = symbols('x y z l')
    Y = Binomial('Y', 2, y)
    Z = Hypergeometric('Z', 10, 2, 3)
    assert factorial_moment(Y, l) == y**2*FallingFactorial(
        2, l) + 2*y*(1 - y)*FallingFactorial(1, l) + (1 - y)**2*\
            FallingFactorial(0, l)
    assert factorial_moment(Z, l) == 7*FallingFactorial(0, l)/\
        15 + 7*FallingFactorial(1, l)/15 + FallingFactorial(2, l)/15


def test_dependence():
    X, Y = Die('X'), Die('Y')
    assert independent(X, 2*Y)
    assert not dependent(X, 2*Y)

    X, Y = Normal('X', 0, 1), Normal('Y', 0, 1)
    assert independent(X, Y)
    assert dependent(X, 2*X)

    # Create a dependency
    XX, YY = given(Tuple(X, Y), Eq(X + Y, 3))
    assert dependent(XX, YY)

def test_dependent_finite():
    X, Y = Die('X'), Die('Y')
    # Dependence testing requires symbolic conditions which currently break
    # finite random variables
    assert dependent(X, Y + X)

    XX, YY = given(Tuple(X, Y), X + Y > 5)  # Create a dependency
    assert dependent(XX, YY)


def test_normality():
    X, Y = Normal('X', 0, 1), Normal('Y', 0, 1)
    x = Symbol('x', real=True)
    z = Symbol('z', real=True)
    dens = density(X - Y, Eq(X + Y, z))

    assert integrate(dens(x), (x, -oo, oo)) == 1


def test_Density():
    X = Die('X', 6)
    d = Density(X)
    assert d.doit() == density(X)

def test_NamedArgsMixin():
    class Foo(Basic, NamedArgsMixin):
        _argnames = 'foo', 'bar'

    a = Foo(S(1), S(2))

    assert a.foo == 1
    assert a.bar == 2

    raises(AttributeError, lambda: a.baz)

    class Bar(Basic, NamedArgsMixin):
        pass

    raises(AttributeError, lambda: Bar(S(1), S(2)).foo)

def test_density_constant():
    assert density(3)(2) == 0
    assert density(3)(3) == DiracDelta(0)

def test_cmoment_constant():
    assert variance(3) == 0
    assert cmoment(3, 3) == 0
    assert cmoment(3, 4) == 0
    x = Symbol('x')
    assert variance(x) == 0
    assert cmoment(x, 15) == 0
    assert cmoment(x, 0) == 1

def test_moment_constant():
    assert moment(3, 0) == 1
    assert moment(3, 1) == 3
    assert moment(3, 2) == 9
    x = Symbol('x')
    assert moment(x, 2) == x**2

def test_median_constant():
    assert median(3) == 3
    x = Symbol('x')
    assert median(x) == x

def test_real():
    x = Normal('x', 0, 1)
    assert x.is_real


def test_issue_10052():
    X = Exponential('X', 3)
    assert P(X < oo) == 1
    assert P(X > oo) == 0
    assert P(X < 2, X > oo) == 0
    assert P(X < oo, X > oo) == 0
    assert P(X < oo, X > 2) == 1
    assert P(X < 3, X == 2) == 0
    raises(ValueError, lambda: P(1))
    raises(ValueError, lambda: P(X < 1, 2))

def test_issue_11934():
    density = {0: .5, 1: .5}
    X = FiniteRV('X', density)
    assert E(X) == 0.5
    assert P( X>= 2) == 0

def test_issue_8129():
    X = Exponential('X', 4)
    assert P(X >= X) == 1
    assert P(X > X) == 0
    assert P(X > X+1) == 0

def test_issue_12237():
    X = Normal('X', 0, 1)
    Y = Normal('Y', 0, 1)
    U = P(X > 0, X)
    V = P(Y < 0, X)
    W = P(X + Y > 0, X)
    assert W == P(X + Y > 0, X)
    assert U == BernoulliDistribution(S.Half, S.Zero, S.One)
    assert V == S.Half

def test_is_random():
    X = Normal('X', 0, 1)
    Y = Normal('Y', 0, 1)
    a, b = symbols('a, b')
    G = GaussianUnitaryEnsemble('U', 2)
    B = BernoulliProcess('B', 0.9)
    assert not is_random(a)
    assert not is_random(a + b)
    assert not is_random(a * b)
    assert not is_random(Matrix([a**2, b**2]))
    assert is_random(X)
    assert is_random(X**2 + Y)
    assert is_random(Y + b**2)
    assert is_random(Y > 5)
    assert is_random(B[3] < 1)
    assert is_random(G)
    assert is_random(X * Y * B[1])
    assert is_random(Matrix([[X, B[2]], [G, Y]]))
    assert is_random(Eq(X, 4))

def test_issue_12283():
    x = symbols('x')
    X = RandomSymbol(x)
    Y = RandomSymbol('Y')
    Z = RandomMatrixSymbol('Z', 2, 1)
    W = RandomMatrixSymbol('W', 2, 1)
    RI = RandomIndexedSymbol(Indexed('RI', 3))
    assert pspace(Z) == PSpace()
    assert pspace(RI) == PSpace()
    assert pspace(X) == PSpace()
    assert E(X) == Expectation(X)
    assert P(Y > 3) == Probability(Y > 3)
    assert variance(X) == Variance(X)
    assert variance(RI) == Variance(RI)
    assert covariance(X, Y) == Covariance(X, Y)
    assert covariance(W, Z) == Covariance(W, Z)

def test_issue_6810():
    X = Die('X', 6)
    Y = Normal('Y', 0, 1)
    assert P(Eq(X, 2)) == S(1)/6
    assert P(Eq(Y, 0)) == 0
    assert P(Or(X > 2, X < 3)) == 1
    assert P(And(X > 3, X > 2)) == S(1)/2

def test_issue_20286():
    n, p = symbols('n p')
    B = Binomial('B', n, p)
    k = Dummy('k', integer = True)
    eq = Sum(Piecewise((-p**k*(1 - p)**(-k + n)*log(p**k*(1 - p)**(-k + n)*binomial(n, k))*binomial(n, k), (k >= 0) & (k <= n)), (nan, True)), (k, 0, n))
    assert eq.dummy_eq(H(B))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\logic\tests\test_lra_theory.py ===
from sympy.core.numbers import Rational, I, oo
from sympy.core.relational import Eq
from sympy.core.symbol import symbols
from sympy.core.singleton import S
from sympy.matrices.dense import Matrix
from sympy.matrices.dense import randMatrix
from sympy.assumptions.ask import Q
from sympy.logic.boolalg import And
from sympy.abc import x, y, z
from sympy.assumptions.cnf import CNF, EncodedCNF
from sympy.functions.elementary.trigonometric import cos
from sympy.external import import_module

from sympy.logic.algorithms.lra_theory import LRASolver, UnhandledInput, LRARational, HANDLE_NEGATION
from sympy.core.random import random, choice, randint
from sympy.core.sympify import sympify
from sympy.ntheory.generate import randprime
from sympy.core.relational import StrictLessThan, StrictGreaterThan
import itertools

from sympy.testing.pytest import raises, XFAIL, skip

def make_random_problem(num_variables=2, num_constraints=2, sparsity=.1, rational=True,
                        disable_strict = False, disable_nonstrict=False, disable_equality=False):
    def rand(sparsity=sparsity):
        if random() < sparsity:
            return sympify(0)
        if rational:
            int1, int2 = [randprime(0, 50) for _ in range(2)]
            return Rational(int1, int2) * choice([-1, 1])
        else:
            return randint(1, 10) * choice([-1, 1])

    variables = symbols('x1:%s' % (num_variables + 1))
    constraints = []
    for _ in range(num_constraints):
        lhs, rhs = sum(rand() * x for x in variables), rand(sparsity=0) # sparsity=0  bc of bug with smtlib_code
        options = []
        if not disable_equality:
            options += [Eq(lhs, rhs)]
        if not disable_nonstrict:
            options += [lhs <= rhs, lhs >= rhs]
        if not disable_strict:
            options += [lhs < rhs, lhs > rhs]

        constraints.append(choice(options))

    return constraints

def check_if_satisfiable_with_z3(constraints):
    from sympy.external.importtools import import_module
    from sympy.printing.smtlib import smtlib_code
    from sympy.logic.boolalg import And
    boolean_formula = And(*constraints)
    z3 = import_module("z3")
    if z3:
        smtlib_string = smtlib_code(boolean_formula)
        s = z3.Solver()
        s.from_string(smtlib_string)
        res = str(s.check())
        if res == 'sat':
            return True
        elif res == 'unsat':
            return False
        else:
            raise ValueError(f"z3 was not able to check the satisfiability of {boolean_formula}")

def find_rational_assignment(constr, assignment, iter=20):
    eps = sympify(1)

    for _ in range(iter):
        assign = {key: val[0] + val[1]*eps for key, val in assignment.items()}
        try:
            for cons in constr:
                assert cons.subs(assign) == True
            return assign
        except AssertionError:
            eps = eps/2

    return None

def boolean_formula_to_encoded_cnf(bf):
    cnf = CNF.from_prop(bf)
    enc = EncodedCNF()
    enc.from_cnf(cnf)
    return enc


def test_from_encoded_cnf():
    s1, s2 = symbols("s1 s2")

    # Test preprocessing
    # Example is from section 3 of paper.
    phi = (x >= 0) & ((x + y <= 2) | (x + 2 * y - z >= 6)) & (Eq(x + y, 2) | (x + 2 * y - z > 4))
    enc = boolean_formula_to_encoded_cnf(phi)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    assert lra.A.shape == (2, 5)
    assert str(lra.slack) == '[_s1, _s2]'
    assert str(lra.nonslack) == '[x, y, z]'
    assert lra.A == Matrix([[ 1,  1, 0, -1,  0],
                            [-1, -2, 1,  0, -1]])
    assert {(str(b.var), b.bound, b.upper, b.equality, b.strict) for b in lra.enc_to_boundary.values()} == {('_s1', 2, None, True, False),
    ('_s1', 2, True, False, False),
    ('_s2', -4, True, False, True),
    ('_s2', -6, True, False, False),
    ('x', 0, False, False, False)}


def test_problem():
    from sympy.logic.algorithms.lra_theory import LRASolver
    from sympy.assumptions.cnf import CNF, EncodedCNF
    cons = [-2 * x - 2 * y >= 7, -9 * y >= 7, -6 * y >= 5]
    cnf = CNF().from_prop(And(*cons))
    enc = EncodedCNF()
    enc.from_cnf(cnf)
    lra, _ = LRASolver.from_encoded_cnf(enc)
    lra.assert_lit(1)
    lra.assert_lit(2)
    lra.assert_lit(3)
    is_sat, assignment = lra.check()
    assert is_sat is True


def test_random_problems():
    z3 = import_module("z3")
    if z3 is None:
        skip("z3 is not installed")

    special_cases = []; x1, x2, x3 = symbols("x1 x2 x3")
    special_cases.append([x1 - 3 * x2 <= -5, 6 * x1 + 4 * x2 <= 0, -7 * x1 + 3 * x2 <= 3])
    special_cases.append([-3 * x1 >= 3, Eq(4 * x1, -1)])
    special_cases.append([-4 * x1 < 4, 6 * x1 <= -6])
    special_cases.append([-3 * x2 >= 7, 6 * x1 <= -5, -3 * x2 <= -4])
    special_cases.append([x + y >= 2, x + y <= 1])
    special_cases.append([x >= 0, x + y <= 2, x + 2 * y - z >= 6])  # from paper example
    special_cases.append([-2 * x1 - 2 * x2 >= 7, -9 * x1 >= 7, -6 * x1 >= 5])
    special_cases.append([2 * x1 > -3, -9 * x1 < -6, 9 * x1 <= 6])
    special_cases.append([-2*x1 < -4, 9*x1 > -9])
    special_cases.append([-6*x1 >= -1, -8*x1 + x2 >= 5, -8*x1 + 7*x2 < 4, x1 > 7])
    special_cases.append([Eq(x1, 2), Eq(5*x1, -2), Eq(-7*x2, -6), Eq(9*x1 + 10*x2, 9)])
    special_cases.append([Eq(3*x1, 6), Eq(x1 - 8*x2, -9), Eq(-7*x1 + 5*x2, 3), Eq(3*x2, 7)])
    special_cases.append([-4*x1 < 4, 6*x1 <= -6])
    special_cases.append([-3*x1 + 8*x2 >= -8, -10*x2 > 9, 8*x1 - 4*x2 < 8, 10*x1 - 9*x2 >= -9])
    special_cases.append([x1 + 5*x2 >= -6, 9*x1 - 3*x2 >= -9, 6*x1 + 6*x2 < -10, -3*x1 + 3*x2 < -7])
    special_cases.append([-9*x1 < 7, -5*x1 - 7*x2 < -1, 3*x1 + 7*x2 > 1, -6*x1 - 6*x2 > 9])
    special_cases.append([9*x1 - 6*x2 >= -7, 9*x1 + 4*x2 < -8, -7*x2 <= 1, 10*x2 <= -7])

    feasible_count = 0
    for i in range(50):
        if i % 8 == 0:
            constraints = make_random_problem(num_variables=1, num_constraints=2, rational=False)
        elif i % 8 == 1:
            constraints = make_random_problem(num_variables=2, num_constraints=4, rational=False, disable_equality=True,
                                              disable_nonstrict=True)
        elif i % 8 == 2:
            constraints = make_random_problem(num_variables=2, num_constraints=4, rational=False, disable_strict=True)
        elif i % 8 == 3:
            constraints = make_random_problem(num_variables=3, num_constraints=12, rational=False)
        else:
            constraints = make_random_problem(num_variables=3, num_constraints=6, rational=False)

        if i < len(special_cases):
            constraints = special_cases[i]

        if False in constraints or True in constraints:
            continue

        phi = And(*constraints)
        if phi == False:
            continue
        cnf = CNF.from_prop(phi); enc = EncodedCNF()
        enc.from_cnf(cnf)
        assert all(0 not in clause for clause in enc.data)

        lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
        s_subs = lra.s_subs

        lra.run_checks = True
        s_subs_rev = {value: key for key, value in s_subs.items()}
        lits = {lit for clause in enc.data for lit in clause}

        bounds = [(lra.enc_to_boundary[l], l) for l in lits if l in lra.enc_to_boundary]
        bounds = sorted(bounds, key=lambda x: (str(x[0].var), x[0].bound, str(x[0].upper))) # to remove nondeterminism

        for b, l in bounds:
            if lra.result and lra.result[0] == False:
                break
            lra.assert_lit(l)

        feasible = lra.check()

        if feasible[0] == True:
            feasible_count += 1
            assert check_if_satisfiable_with_z3(constraints) is True
            cons_funcs = [cons.func for cons in constraints]
            assignment = feasible[1]
            assignment = {key.var : value for key, value in assignment.items()}
            if not (StrictLessThan in cons_funcs or StrictGreaterThan in cons_funcs):
                assignment = {key: value[0] for key, value in assignment.items()}
                for cons in constraints:
                    assert cons.subs(assignment) == True

            else:
                rat_assignment = find_rational_assignment(constraints, assignment)
                assert rat_assignment is not None
        else:
            assert check_if_satisfiable_with_z3(constraints) is False

            conflict = feasible[1]
            assert len(conflict) >= 2
            conflict = {lra.enc_to_boundary[-l].get_inequality() for l in conflict}
            conflict = {clause.subs(s_subs_rev) for clause in conflict}
            assert check_if_satisfiable_with_z3(conflict) is False

            # check that conflict clause is probably minimal
            for subset in itertools.combinations(conflict, len(conflict)-1):
                assert check_if_satisfiable_with_z3(subset) is True


@XFAIL
def test_pos_neg_zero():
    bf = Q.positive(x) & Q.negative(x) & Q.zero(y)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for lit in enc.encoding.values():
        if lra.assert_lit(lit) is not None:
            break
    assert len(lra.enc_to_boundary) == 3
    assert lra.check()[0] == False

    bf = Q.positive(x) & Q.lt(x, -1)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for lit in enc.encoding.values():
        if lra.assert_lit(lit) is not None:
            break
    assert len(lra.enc_to_boundary) == 2
    assert lra.check()[0] == False

    bf = Q.positive(x) & Q.zero(x)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for lit in enc.encoding.values():
        if lra.assert_lit(lit) is not None:
            break
    assert len(lra.enc_to_boundary) == 2
    assert lra.check()[0] == False

    bf = Q.positive(x) & Q.zero(y)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for lit in enc.encoding.values():
        if lra.assert_lit(lit) is not None:
            break
    assert len(lra.enc_to_boundary) == 2
    assert lra.check()[0] == True


@XFAIL
def test_pos_neg_infinite():
    bf = Q.positive_infinite(x) & Q.lt(x, 10000000) & Q.positive_infinite(y)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for lit in enc.encoding.values():
        if lra.assert_lit(lit) is not None:
            break
    assert len(lra.enc_to_boundary) == 3
    assert lra.check()[0] == False

    bf = Q.positive_infinite(x) & Q.gt(x, 10000000) & Q.positive_infinite(y)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for lit in enc.encoding.values():
        if lra.assert_lit(lit) is not None:
            break
    assert len(lra.enc_to_boundary) == 3
    assert lra.check()[0] == True

    bf = Q.positive_infinite(x) & Q.negative_infinite(x)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for lit in enc.encoding.values():
        if lra.assert_lit(lit) is not None:
            break
    assert len(lra.enc_to_boundary) == 2
    assert lra.check()[0] == False


def test_binrel_evaluation():
    bf = Q.gt(3, 2)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, conflicts = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    assert len(lra.enc_to_boundary) == 0
    assert conflicts == [[1]]

    bf = Q.lt(3, 2)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, conflicts = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    assert len(lra.enc_to_boundary) == 0
    assert conflicts == [[-1]]


def test_negation():
    assert HANDLE_NEGATION is True
    bf = Q.gt(x, 1) & ~Q.gt(x, 0)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for clause in enc.data:
        for lit in clause:
            lra.assert_lit(lit)
    assert len(lra.enc_to_boundary) == 2
    assert lra.check()[0] == False
    assert sorted(lra.check()[1]) in [[-1, 2], [-2, 1]]

    bf = ~Q.gt(x, 1) & ~Q.lt(x, 0)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for clause in enc.data:
        for lit in clause:
            lra.assert_lit(lit)
    assert len(lra.enc_to_boundary) == 2
    assert lra.check()[0] == True

    bf = ~Q.gt(x, 0) & ~Q.lt(x, 1)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for clause in enc.data:
        for lit in clause:
            lra.assert_lit(lit)
    assert len(lra.enc_to_boundary) == 2
    assert lra.check()[0] == False

    bf = ~Q.gt(x, 0) & ~Q.le(x, 0)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for clause in enc.data:
        for lit in clause:
            lra.assert_lit(lit)
    assert len(lra.enc_to_boundary) == 2
    assert lra.check()[0] == False

    bf = ~Q.le(x+y, 2) & ~Q.ge(x-y, 2) & ~Q.ge(y, 0)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for clause in enc.data:
        for lit in clause:
            lra.assert_lit(lit)
    assert len(lra.enc_to_boundary) == 3
    assert lra.check()[0] == False
    assert len(lra.check()[1]) == 3
    assert all(i > 0 for i in lra.check()[1])


def test_unhandled_input():
    nan = S.NaN
    bf = Q.gt(3, nan) & Q.gt(x, nan)
    enc = boolean_formula_to_encoded_cnf(bf)
    raises(ValueError, lambda: LRASolver.from_encoded_cnf(enc, testing_mode=True))

    bf = Q.gt(3, I) & Q.gt(x, I)
    enc = boolean_formula_to_encoded_cnf(bf)
    raises(UnhandledInput, lambda: LRASolver.from_encoded_cnf(enc, testing_mode=True))

    bf = Q.gt(3, float("inf")) & Q.gt(x, float("inf"))
    enc = boolean_formula_to_encoded_cnf(bf)
    raises(UnhandledInput, lambda: LRASolver.from_encoded_cnf(enc, testing_mode=True))

    bf = Q.gt(3, oo) & Q.gt(x, oo)
    enc = boolean_formula_to_encoded_cnf(bf)
    raises(UnhandledInput, lambda: LRASolver.from_encoded_cnf(enc, testing_mode=True))

    # test non-linearity
    bf = Q.gt(x**2 + x, 2)
    enc = boolean_formula_to_encoded_cnf(bf)
    raises(UnhandledInput, lambda: LRASolver.from_encoded_cnf(enc, testing_mode=True))

    bf = Q.gt(cos(x) + x, 2)
    enc = boolean_formula_to_encoded_cnf(bf)
    raises(UnhandledInput, lambda: LRASolver.from_encoded_cnf(enc, testing_mode=True))

@XFAIL
def test_infinite_strict_inequalities():
    # Extensive testing of the interaction between strict inequalities
    # and constraints containing infinity is needed because
    # the paper's rule for strict inequalities don't work when
    # infinite numbers are allowed. Using the paper's rules you
    # can end up with situations where oo + delta > oo is considered
    # True when oo + delta should be equal to oo.
    # See https://math.stackexchange.com/questions/4757069/can-this-method-of-converting-strict-inequalities-to-equisatisfiable-nonstrict-i
    bf = (-x - y >= -float("inf")) & (x > 0) & (y >= float("inf"))
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for lit in sorted(enc.encoding.values()):
        if lra.assert_lit(lit) is not None:
            break
    assert len(lra.enc_to_boundary) == 3
    assert lra.check()[0] == True


def test_pivot():
    for _ in range(10):
        m = randMatrix(5)
        rref = m.rref()
        for _ in range(5):
            i, j = randint(0, 4), randint(0, 4)
            if m[i, j] != 0:
                assert LRASolver._pivot(m, i, j).rref() == rref


def test_reset_bounds():
    bf = Q.ge(x, 1) & Q.lt(x, 1)
    enc = boolean_formula_to_encoded_cnf(bf)
    lra, _ = LRASolver.from_encoded_cnf(enc, testing_mode=True)
    for clause in enc.data:
        for lit in clause:
            lra.assert_lit(lit)
    assert len(lra.enc_to_boundary) == 2
    assert lra.check()[0] == False

    lra.reset_bounds()
    assert lra.check()[0] == True
    for var in lra.all_var:
        assert var.upper == LRARational(float("inf"), 0)
        assert var.upper_from_eq == False
        assert var.upper_from_neg == False
        assert var.lower == LRARational(-float("inf"), 0)
        assert var.lower_from_eq == False
        assert var.lower_from_neg == False
        assert var.assign == LRARational(0, 0)
        assert var.var is not None
        assert var.col_idx is not None


def test_empty_cnf():
    cnf = CNF()
    enc = EncodedCNF()
    enc.from_cnf(cnf)
    lra, conflict = LRASolver.from_encoded_cnf(enc)
    assert len(conflict) == 0
    assert lra.check() == (True, {})

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\bs4\tests\test_pageelement.py ===
"""Tests of the bs4.element.PageElement class"""

import copy
import pickle
import pytest
import sys
import warnings

from bs4 import BeautifulSoup
from bs4.element import (
    AttributeValueList,
    Comment,
)
from bs4.filter import SoupStrainer
from . import (
    SoupTest,
)


class TestEncoding(SoupTest):
    """Test the ability to encode objects into strings."""

    def test_unicode_string_can_be_encoded(self):
        html = "<b>\N{SNOWMAN}</b>"
        soup = self.soup(html)
        assert soup.b.string.encode("utf-8") == "\N{SNOWMAN}".encode("utf-8")

    def test_tag_containing_unicode_string_can_be_encoded(self):
        html = "<b>\N{SNOWMAN}</b>"
        soup = self.soup(html)
        assert soup.b.encode("utf-8") == html.encode("utf-8")

    def test_encoding_substitutes_unrecognized_characters_by_default(self):
        html = "<b>\N{SNOWMAN}</b>"
        soup = self.soup(html)
        assert soup.b.encode("ascii") == b"<b>&#9731;</b>"

    def test_encoding_can_be_made_strict(self):
        html = "<b>\N{SNOWMAN}</b>"
        soup = self.soup(html)
        with pytest.raises(UnicodeEncodeError):
            soup.encode("ascii", errors="strict")

    def test_decode_contents(self):
        html = "<b>\N{SNOWMAN}</b>"
        soup = self.soup(html)
        assert "\N{SNOWMAN}" == soup.b.decode_contents()

    def test_encode_contents(self):
        html = "<b>\N{SNOWMAN}</b>"
        soup = self.soup(html)
        assert "\N{SNOWMAN}".encode("utf8") == soup.b.encode_contents(encoding="utf8")

    def test_encode_deeply_nested_document(self):
        # This test verifies that encoding a string doesn't involve
        # any recursive function calls. If it did, this test would
        # overflow the Python interpreter stack.
        limit = sys.getrecursionlimit() + 1
        markup = "<span>" * limit
        soup = self.soup(markup)
        encoded = soup.encode()
        assert limit == encoded.count(b"<span>")

    def test_deprecated_renderContents(self):
        html = "<b>\N{SNOWMAN}</b>"
        soup = self.soup(html)
        with warnings.catch_warnings(record=True) as w:
            soup.renderContents()
            assert "\N{SNOWMAN}".encode("utf8") == soup.b.renderContents()
        msgs = [str(warning.message) for warning in w]
        assert all(
            x
            == "Call to deprecated method renderContents. (Replaced by encode_contents) -- Deprecated since version 4.0.0."
            for x in msgs
        )

    def test_repr(self):
        html = "<b>\N{SNOWMAN}</b>"
        soup = self.soup(html)
        assert html == repr(soup)


class TestFormatters(SoupTest):
    """Test the formatting feature, used by methods like decode() and
    prettify(), and the formatters themselves.
    """

    def test_default_formatter_is_minimal(self):
        markup = "<b>&lt;&lt;Sacr\N{LATIN SMALL LETTER E WITH ACUTE} bleu!&gt;&gt;</b>"
        soup = self.soup(markup)
        decoded = soup.decode(formatter="minimal")
        # The < is converted back into &lt; but the e-with-acute is left alone.
        assert decoded == self.document_for(
            "<b>&lt;&lt;Sacr\N{LATIN SMALL LETTER E WITH ACUTE} bleu!&gt;&gt;</b>"
        )

    def test_formatter_html(self):
        markup = (
            "<br><b>&lt;&lt;Sacr\N{LATIN SMALL LETTER E WITH ACUTE} bleu!&gt;&gt;</b>"
        )
        soup = self.soup(markup)
        decoded = soup.decode(formatter="html")
        assert decoded == self.document_for(
            "<br/><b>&lt;&lt;Sacr&eacute; bleu!&gt;&gt;</b>"
        )

    def test_formatter_html5(self):
        markup = (
            "<br><b>&lt;&lt;Sacr\N{LATIN SMALL LETTER E WITH ACUTE} bleu!&gt;&gt;</b>"
        )
        soup = self.soup(markup)
        decoded = soup.decode(formatter="html5")
        assert decoded == self.document_for(
            "<br><b>&lt;&lt;Sacr&eacute; bleu!&gt;&gt;</b>"
        )

    def test_formatter_minimal(self):
        markup = "<b>&lt;&lt;Sacr\N{LATIN SMALL LETTER E WITH ACUTE} bleu!&gt;&gt;</b>"
        soup = self.soup(markup)
        decoded = soup.decode(formatter="minimal")
        # The < is converted back into &lt; but the e-with-acute is left alone.
        assert decoded == self.document_for(
            "<b>&lt;&lt;Sacr\N{LATIN SMALL LETTER E WITH ACUTE} bleu!&gt;&gt;</b>"
        )

    def test_formatter_null(self):
        markup = "<b>&lt;&lt;Sacr\N{LATIN SMALL LETTER E WITH ACUTE} bleu!&gt;&gt;</b>"
        soup = self.soup(markup)
        decoded = soup.decode(formatter=None)
        # Neither the angle brackets nor the e-with-acute are converted.
        # This is not valid HTML, but it's what the user wanted.
        assert decoded == self.document_for(
            "<b><<Sacr\N{LATIN SMALL LETTER E WITH ACUTE} bleu!>></b>"
        )

    def test_formatter_custom(self):
        markup = "<b>&lt;foo&gt;</b><b>bar</b><br/>"
        soup = self.soup(markup)
        decoded = soup.decode(formatter=lambda x: x.upper())
        # Instead of normal entity conversion code, the custom
        # callable is called on every string.
        assert decoded == self.document_for("<b><FOO></b><b>BAR</b><br/>")

    def test_formatter_is_run_on_attribute_values(self):
        markup = '<a href="http://a.com?a=b&c=é">e</a>'
        soup = self.soup(markup)
        a = soup.a

        expect_minimal = '<a href="http://a.com?a=b&amp;c=é">e</a>'

        assert expect_minimal == a.decode()
        assert expect_minimal == a.decode(formatter="minimal")

        expect_html = '<a href="http://a.com?a=b&amp;c=&eacute;">e</a>'
        assert expect_html == a.decode(formatter="html")

        assert markup == a.decode(formatter=None)
        expect_upper = '<a href="HTTP://A.COM?A=B&C=É">E</a>'
        assert expect_upper == a.decode(formatter=lambda x: x.upper())

    def test_formatter_skips_script_tag_for_html_documents(self):
        doc = """
  <script type="text/javascript">
   console.log("< < hey > > ");
  </script>
"""
        encoded = BeautifulSoup(doc, "html.parser").encode()
        assert b"< < hey > >" in encoded

    def test_formatter_skips_style_tag_for_html_documents(self):
        doc = """
  <style type="text/css">
   console.log("< < hey > > ");
  </style>
"""
        encoded = BeautifulSoup(doc, "html.parser").encode()
        assert b"< < hey > >" in encoded

    def test_prettify_leaves_preformatted_text_alone(self):
        soup = self.soup(
            "<div>  foo  <pre>  \tbar\n  \n  </pre>  baz  <textarea> eee\nfff\t</textarea></div>"
        )
        # Everything outside the <pre> tag is reformatted, but everything
        # inside is left alone.
        assert (
            "<div>\n foo\n <pre>  \tbar\n  \n  </pre>\n baz\n <textarea> eee\nfff\t</textarea>\n</div>\n"
            == soup.div.prettify()
        )

    def test_prettify_handles_nested_string_literal_tags(self):
        # Most of this markup is inside a <pre> tag, so prettify()
        # only does three things to it:
        # 1. Add a newline and a space between the <div> and the <pre>
        # 2. Add a newline after the </pre>
        # 3. Add a newline at the end.
        #
        # The contents of the <pre> tag are left completely alone.  In
        # particular, we don't start adding whitespace again once we
        # encounter the first </pre> tag, because we know it's not
        # the one that put us into string literal mode.
        markup = """<div><pre><code>some
<script><pre>code</pre></script> for you
</code></pre></div>"""

        expect = """<div>
 <pre><code>some
<script><pre>code</pre></script> for you
</code></pre>
</div>
"""
        soup = self.soup(markup)
        assert expect == soup.div.prettify()

    def test_prettify_accepts_formatter_function(self):
        soup = BeautifulSoup("<html><body>foo</body></html>", "html.parser")
        pretty = soup.prettify(formatter=lambda x: x.upper())
        assert "FOO" in pretty

    def test_prettify_outputs_unicode_by_default(self):
        soup = self.soup("<a></a>")
        assert str is type(soup.prettify())

    def test_prettify_can_encode_data(self):
        soup = self.soup("<a></a>")
        assert bytes is type(soup.prettify("utf-8"))

    def test_html_entity_substitution_off_by_default(self):
        markup = "<b>Sacr\N{LATIN SMALL LETTER E WITH ACUTE} bleu!</b>"
        soup = self.soup(markup)
        encoded = soup.b.encode("utf-8")
        assert encoded == markup.encode("utf-8")

    def test_encoding_substitution(self):
        # Here's the <meta> tag saying that a document is
        # encoded in Shift-JIS.
        meta_tag = (
            '<meta content="text/html; charset=x-sjis" ' 'http-equiv="Content-type"/>'
        )
        soup = self.soup(meta_tag)

        # Parse the document, and the charset apprears unchanged.
        assert soup.meta["content"] == "text/html; charset=x-sjis"

        # Encode the document into some encoding, and the encoding is
        # substituted into the meta tag.
        utf_8 = soup.encode("utf-8")
        assert b"charset=utf-8" in utf_8

        euc_jp = soup.encode("euc_jp")
        assert b"charset=euc_jp" in euc_jp

        shift_jis = soup.encode("shift-jis")
        assert b"charset=shift-jis" in shift_jis

        utf_16_u = soup.encode("utf-16").decode("utf-16")
        assert "charset=utf-16" in utf_16_u

    def test_encoding_substitution_doesnt_happen_if_tag_is_strained(self):
        markup = (
            '<head><meta content="text/html; charset=x-sjis" '
            'http-equiv="Content-type"/></head><pre>foo</pre>'
        )

        # Beautiful Soup used to try to rewrite the meta tag even if the
        # meta tag got filtered out by the strainer. This test makes
        # sure that doesn't happen.
        strainer = SoupStrainer("pre")
        soup = self.soup(markup, parse_only=strainer)
        assert soup.contents[0].name == "pre"


class TestPersistence(SoupTest):
    "Testing features like pickle and deepcopy."

    def setup_method(self):
        self.page = """<!DOCTYPE HTML PUBLIC "-//W3C//DTD HTML 4.0 Transitional//EN"
"http://www.w3.org/TR/REC-html40/transitional.dtd">
<html>
<head>
<meta http-equiv="Content-Type" content="text/html; charset=utf-8">
<title>Beautiful Soup: We called him Tortoise because he taught us.</title>
<link rev="made" href="mailto:leonardr@segfault.org">
<meta name="Description" content="Beautiful Soup: an HTML parser optimized for screen-scraping.">
<meta name="generator" content="Markov Approximation 1.4 (module: leonardr)">
<meta name="author" content="Leonard Richardson">
</head>
<body>
<a href="foo">foo</a>
<a href="foo"><b>bar</b></a>
</body>
</html>"""
        self.tree = self.soup(self.page)

    def test_pickle_and_unpickle_identity(self):
        # Pickling a tree, then unpickling it, yields a tree identical
        # to the original.
        dumped = pickle.dumps(self.tree, 2)
        loaded = pickle.loads(dumped)
        assert loaded.__class__ == BeautifulSoup
        assert loaded.decode() == self.tree.decode()

    def test_deepcopy_identity(self):
        # Making a deepcopy of a tree yields an identical tree.
        copied = copy.deepcopy(self.tree)
        assert copied.decode() == self.tree.decode()

    def test_copy_deeply_nested_document(self):
        # This test verifies that copy and deepcopy don't involve any
        # recursive function calls. If they did, this test would
        # overflow the Python interpreter stack.
        limit = sys.getrecursionlimit() + 1
        markup = "<span>" * limit

        soup = self.soup(markup)

        copy.copy(soup)
        copy.deepcopy(soup)

    def test_copy_preserves_encoding(self):
        soup = BeautifulSoup(b"<p>&nbsp;</p>", "html.parser")
        encoding = soup.original_encoding
        copy = soup.__copy__()
        assert "<p> </p>" == str(copy)
        assert encoding == copy.original_encoding

    def test_copy_preserves_builder_information(self):
        tag = self.soup("<p></p>").p

        # Simulate a tag obtained from a source file.
        tag.sourceline = 10
        tag.sourcepos = 33

        copied = tag.__copy__()

        # The TreeBuilder object is no longer availble, but information
        # obtained from it gets copied over to the new Tag object.
        assert tag.sourceline == copied.sourceline
        assert tag.sourcepos == copied.sourcepos
        assert tag.can_be_empty_element == copied.can_be_empty_element
        assert tag.cdata_list_attributes == copied.cdata_list_attributes
        assert tag.preserve_whitespace_tags == copied.preserve_whitespace_tags
        assert tag.interesting_string_types == copied.interesting_string_types

    def test_unicode_pickle(self):
        # A tree containing Unicode characters can be pickled.
        html = "<b>\N{SNOWMAN}</b>"
        soup = self.soup(html)
        dumped = pickle.dumps(soup, pickle.HIGHEST_PROTOCOL)
        loaded = pickle.loads(dumped)
        assert loaded.decode() == soup.decode()

    def test_copy_navigablestring_is_not_attached_to_tree(self):
        html = "<b>Foo<a></a></b><b>Bar</b>"
        soup = self.soup(html)
        s1 = soup.find(string="Foo")
        s2 = copy.copy(s1)
        assert s1 == s2
        assert None is s2.parent
        assert None is s2.next_element
        assert None is not s1.next_sibling
        assert None is s2.next_sibling
        assert None is s2.previous_element

    def test_copy_navigablestring_subclass_has_same_type(self):
        html = "<b><!--Foo--></b>"
        soup = self.soup(html)
        s1 = soup.string
        s2 = copy.copy(s1)
        assert s1 == s2
        assert isinstance(s2, Comment)

    def test_copy_entire_soup(self):
        html = "<div><b>Foo<a></a></b><b>Bar</b></div>end"
        soup = self.soup(html)
        soup_copy = copy.copy(soup)
        assert soup == soup_copy

    def test_copy_tag_copies_contents(self):
        html = "<div class='a b c'><b>Foo<a></a></b><b>Bar</b></div>end"
        soup = self.soup(html)
        div = soup.div
        div_copy = copy.copy(div)

        # The two tags look the same, and evaluate to equal.
        assert str(div) == str(div_copy)
        assert div == div_copy

        # But they're not the same object.
        assert div is not div_copy

        # And they don't have the same relation to the parse tree. The
        # copy is not associated with a parse tree at all.
        assert None is div_copy.parent
        assert None is div_copy.previous_element
        assert None is div_copy.find(string="Bar").next_element
        assert None is not div.find(string="Bar").next_element

        # Modifying one of the tag's multi-valued attributes
        # doesn't modify the other.
        assert div["class"] is not div_copy["class"]
        div["class"].append("d")
        assert "a b c d".split() == div["class"]
        assert "a b c".split() == div_copy["class"]
        assert isinstance(div_copy["class"], AttributeValueList)


class TestEquality(SoupTest):

    def test_comparison(self):
        soup = self.soup("<a>string</a> <a>string</a>")
        first_a, second_a = soup.find_all('a')
        first_string, second_string = soup.find_all(string='string')

        # Tags with the same markup are equal.
        assert first_a == second_a

        # NavigableStrings with the same content are equal, and also
        # equal to a Python string with the same content...
        assert first_string == second_string == "string"

        # ...but not equivalent to a bytestring with the same content.
        assert first_string != b"string"

    def test_hash(self):
        soup = self.soup("<a>string</a> <a>string</a>")
        first_a, second_a = soup.find_all('a')
        first_string, second_string = soup.find_all(string='string')

        # Tags with the same markup hash to the same value.
        assert hash(first_a) == hash(second_a)

        # But they're not the same object.
        assert id(first_a) != id(second_a)

        # NavigableStrings with the same contents hash to the value of
        # the contents.
        assert hash(first_string) == hash(second_string) == hash("string")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\config\test_config.py ===
import pytest

from pandas._config import config as cf
from pandas._config.config import OptionError

import pandas as pd
import pandas._testing as tm


class TestConfig:
    @pytest.fixture(autouse=True)
    def clean_config(self, monkeypatch):
        with monkeypatch.context() as m:
            m.setattr(cf, "_global_config", {})
            m.setattr(cf, "options", cf.DictWrapper(cf._global_config))
            m.setattr(cf, "_deprecated_options", {})
            m.setattr(cf, "_registered_options", {})

            # Our test fixture in conftest.py sets "chained_assignment"
            # to "raise" only after all test methods have been setup.
            # However, after this setup, there is no longer any
            # "chained_assignment" option, so re-register it.
            cf.register_option("chained_assignment", "raise")
            yield

    def test_api(self):
        # the pandas object exposes the user API
        assert hasattr(pd, "get_option")
        assert hasattr(pd, "set_option")
        assert hasattr(pd, "reset_option")
        assert hasattr(pd, "describe_option")

    def test_is_one_of_factory(self):
        v = cf.is_one_of_factory([None, 12])

        v(12)
        v(None)
        msg = r"Value must be one of None\|12"
        with pytest.raises(ValueError, match=msg):
            v(1.1)

    def test_register_option(self):
        cf.register_option("a", 1, "doc")

        # can't register an already registered option
        msg = "Option 'a' has already been registered"
        with pytest.raises(OptionError, match=msg):
            cf.register_option("a", 1, "doc")

        # can't register an already registered option
        msg = "Path prefix to option 'a' is already an option"
        with pytest.raises(OptionError, match=msg):
            cf.register_option("a.b.c.d1", 1, "doc")
        with pytest.raises(OptionError, match=msg):
            cf.register_option("a.b.c.d2", 1, "doc")

        # no python keywords
        msg = "for is a python keyword"
        with pytest.raises(ValueError, match=msg):
            cf.register_option("for", 0)
        with pytest.raises(ValueError, match=msg):
            cf.register_option("a.for.b", 0)
        # must be valid identifier (ensure attribute access works)
        msg = "oh my goddess! is not a valid identifier"
        with pytest.raises(ValueError, match=msg):
            cf.register_option("Oh my Goddess!", 0)

        # we can register options several levels deep
        # without predefining the intermediate steps
        # and we can define differently named options
        # in the same namespace
        cf.register_option("k.b.c.d1", 1, "doc")
        cf.register_option("k.b.c.d2", 1, "doc")

    def test_describe_option(self):
        cf.register_option("a", 1, "doc")
        cf.register_option("b", 1, "doc2")
        cf.deprecate_option("b")

        cf.register_option("c.d.e1", 1, "doc3")
        cf.register_option("c.d.e2", 1, "doc4")
        cf.register_option("f", 1)
        cf.register_option("g.h", 1)
        cf.register_option("k", 2)
        cf.deprecate_option("g.h", rkey="k")
        cf.register_option("l", "foo")

        # non-existent keys raise KeyError
        msg = r"No such keys\(s\)"
        with pytest.raises(OptionError, match=msg):
            cf.describe_option("no.such.key")

        # we can get the description for any key we registered
        assert "doc" in cf.describe_option("a", _print_desc=False)
        assert "doc2" in cf.describe_option("b", _print_desc=False)
        assert "precated" in cf.describe_option("b", _print_desc=False)
        assert "doc3" in cf.describe_option("c.d.e1", _print_desc=False)
        assert "doc4" in cf.describe_option("c.d.e2", _print_desc=False)

        # if no doc is specified we get a default message
        # saying "description not available"
        assert "available" in cf.describe_option("f", _print_desc=False)
        assert "available" in cf.describe_option("g.h", _print_desc=False)
        assert "precated" in cf.describe_option("g.h", _print_desc=False)
        assert "k" in cf.describe_option("g.h", _print_desc=False)

        # default is reported
        assert "foo" in cf.describe_option("l", _print_desc=False)
        # current value is reported
        assert "bar" not in cf.describe_option("l", _print_desc=False)
        cf.set_option("l", "bar")
        assert "bar" in cf.describe_option("l", _print_desc=False)

    def test_case_insensitive(self):
        cf.register_option("KanBAN", 1, "doc")

        assert "doc" in cf.describe_option("kanbaN", _print_desc=False)
        assert cf.get_option("kanBaN") == 1
        cf.set_option("KanBan", 2)
        assert cf.get_option("kAnBaN") == 2

        # gets of non-existent keys fail
        msg = r"No such keys\(s\): 'no_such_option'"
        with pytest.raises(OptionError, match=msg):
            cf.get_option("no_such_option")
        cf.deprecate_option("KanBan")

        assert cf._is_deprecated("kAnBaN")

    def test_get_option(self):
        cf.register_option("a", 1, "doc")
        cf.register_option("b.c", "hullo", "doc2")
        cf.register_option("b.b", None, "doc2")

        # gets of existing keys succeed
        assert cf.get_option("a") == 1
        assert cf.get_option("b.c") == "hullo"
        assert cf.get_option("b.b") is None

        # gets of non-existent keys fail
        msg = r"No such keys\(s\): 'no_such_option'"
        with pytest.raises(OptionError, match=msg):
            cf.get_option("no_such_option")

    def test_set_option(self):
        cf.register_option("a", 1, "doc")
        cf.register_option("b.c", "hullo", "doc2")
        cf.register_option("b.b", None, "doc2")

        assert cf.get_option("a") == 1
        assert cf.get_option("b.c") == "hullo"
        assert cf.get_option("b.b") is None

        cf.set_option("a", 2)
        cf.set_option("b.c", "wurld")
        cf.set_option("b.b", 1.1)

        assert cf.get_option("a") == 2
        assert cf.get_option("b.c") == "wurld"
        assert cf.get_option("b.b") == 1.1

        msg = r"No such keys\(s\): 'no.such.key'"
        with pytest.raises(OptionError, match=msg):
            cf.set_option("no.such.key", None)

    def test_set_option_empty_args(self):
        msg = "Must provide an even number of non-keyword arguments"
        with pytest.raises(ValueError, match=msg):
            cf.set_option()

    def test_set_option_uneven_args(self):
        msg = "Must provide an even number of non-keyword arguments"
        with pytest.raises(ValueError, match=msg):
            cf.set_option("a.b", 2, "b.c")

    def test_set_option_invalid_single_argument_type(self):
        msg = "Must provide an even number of non-keyword arguments"
        with pytest.raises(ValueError, match=msg):
            cf.set_option(2)

    def test_set_option_multiple(self):
        cf.register_option("a", 1, "doc")
        cf.register_option("b.c", "hullo", "doc2")
        cf.register_option("b.b", None, "doc2")

        assert cf.get_option("a") == 1
        assert cf.get_option("b.c") == "hullo"
        assert cf.get_option("b.b") is None

        cf.set_option("a", "2", "b.c", None, "b.b", 10.0)

        assert cf.get_option("a") == "2"
        assert cf.get_option("b.c") is None
        assert cf.get_option("b.b") == 10.0

    def test_validation(self):
        cf.register_option("a", 1, "doc", validator=cf.is_int)
        cf.register_option("d", 1, "doc", validator=cf.is_nonnegative_int)
        cf.register_option("b.c", "hullo", "doc2", validator=cf.is_text)

        msg = "Value must have type '<class 'int'>'"
        with pytest.raises(ValueError, match=msg):
            cf.register_option("a.b.c.d2", "NO", "doc", validator=cf.is_int)

        cf.set_option("a", 2)  # int is_int
        cf.set_option("b.c", "wurld")  # str is_str
        cf.set_option("d", 2)
        cf.set_option("d", None)  # non-negative int can be None

        # None not is_int
        with pytest.raises(ValueError, match=msg):
            cf.set_option("a", None)
        with pytest.raises(ValueError, match=msg):
            cf.set_option("a", "ab")

        msg = "Value must be a nonnegative integer or None"
        with pytest.raises(ValueError, match=msg):
            cf.register_option("a.b.c.d3", "NO", "doc", validator=cf.is_nonnegative_int)
        with pytest.raises(ValueError, match=msg):
            cf.register_option("a.b.c.d3", -2, "doc", validator=cf.is_nonnegative_int)

        msg = r"Value must be an instance of <class 'str'>\|<class 'bytes'>"
        with pytest.raises(ValueError, match=msg):
            cf.set_option("b.c", 1)

        validator = cf.is_one_of_factory([None, cf.is_callable])
        cf.register_option("b", lambda: None, "doc", validator=validator)
        # pylint: disable-next=consider-using-f-string
        cf.set_option("b", "%.1f".format)  # Formatter is callable
        cf.set_option("b", None)  # Formatter is none (default)
        with pytest.raises(ValueError, match="Value must be a callable"):
            cf.set_option("b", "%.1f")

    def test_reset_option(self):
        cf.register_option("a", 1, "doc", validator=cf.is_int)
        cf.register_option("b.c", "hullo", "doc2", validator=cf.is_str)
        assert cf.get_option("a") == 1
        assert cf.get_option("b.c") == "hullo"

        cf.set_option("a", 2)
        cf.set_option("b.c", "wurld")
        assert cf.get_option("a") == 2
        assert cf.get_option("b.c") == "wurld"

        cf.reset_option("a")
        assert cf.get_option("a") == 1
        assert cf.get_option("b.c") == "wurld"
        cf.reset_option("b.c")
        assert cf.get_option("a") == 1
        assert cf.get_option("b.c") == "hullo"

    def test_reset_option_all(self):
        cf.register_option("a", 1, "doc", validator=cf.is_int)
        cf.register_option("b.c", "hullo", "doc2", validator=cf.is_str)
        assert cf.get_option("a") == 1
        assert cf.get_option("b.c") == "hullo"

        cf.set_option("a", 2)
        cf.set_option("b.c", "wurld")
        assert cf.get_option("a") == 2
        assert cf.get_option("b.c") == "wurld"

        cf.reset_option("all")
        assert cf.get_option("a") == 1
        assert cf.get_option("b.c") == "hullo"

    def test_deprecate_option(self):
        # we can deprecate non-existent options
        cf.deprecate_option("foo")

        assert cf._is_deprecated("foo")
        with tm.assert_produces_warning(FutureWarning, match="deprecated"):
            with pytest.raises(KeyError, match="No such keys.s.: 'foo'"):
                cf.get_option("foo")

        cf.register_option("a", 1, "doc", validator=cf.is_int)
        cf.register_option("b.c", "hullo", "doc2")
        cf.register_option("foo", "hullo", "doc2")

        cf.deprecate_option("a", removal_ver="nifty_ver")
        with tm.assert_produces_warning(FutureWarning, match="eprecated.*nifty_ver"):
            cf.get_option("a")

            msg = "Option 'a' has already been defined as deprecated"
            with pytest.raises(OptionError, match=msg):
                cf.deprecate_option("a")

        cf.deprecate_option("b.c", "zounds!")
        with tm.assert_produces_warning(FutureWarning, match="zounds!"):
            cf.get_option("b.c")

        # test rerouting keys
        cf.register_option("d.a", "foo", "doc2")
        cf.register_option("d.dep", "bar", "doc2")
        assert cf.get_option("d.a") == "foo"
        assert cf.get_option("d.dep") == "bar"

        cf.deprecate_option("d.dep", rkey="d.a")  # reroute d.dep to d.a
        with tm.assert_produces_warning(FutureWarning, match="eprecated"):
            assert cf.get_option("d.dep") == "foo"

        with tm.assert_produces_warning(FutureWarning, match="eprecated"):
            cf.set_option("d.dep", "baz")  # should overwrite "d.a"

        with tm.assert_produces_warning(FutureWarning, match="eprecated"):
            assert cf.get_option("d.dep") == "baz"

    def test_config_prefix(self):
        with cf.config_prefix("base"):
            cf.register_option("a", 1, "doc1")
            cf.register_option("b", 2, "doc2")
            assert cf.get_option("a") == 1
            assert cf.get_option("b") == 2

            cf.set_option("a", 3)
            cf.set_option("b", 4)
            assert cf.get_option("a") == 3
            assert cf.get_option("b") == 4

        assert cf.get_option("base.a") == 3
        assert cf.get_option("base.b") == 4
        assert "doc1" in cf.describe_option("base.a", _print_desc=False)
        assert "doc2" in cf.describe_option("base.b", _print_desc=False)

        cf.reset_option("base.a")
        cf.reset_option("base.b")

        with cf.config_prefix("base"):
            assert cf.get_option("a") == 1
            assert cf.get_option("b") == 2

    def test_callback(self):
        k = [None]
        v = [None]

        def callback(key):
            k.append(key)
            v.append(cf.get_option(key))

        cf.register_option("d.a", "foo", cb=callback)
        cf.register_option("d.b", "foo", cb=callback)

        del k[-1], v[-1]
        cf.set_option("d.a", "fooz")
        assert k[-1] == "d.a"
        assert v[-1] == "fooz"

        del k[-1], v[-1]
        cf.set_option("d.b", "boo")
        assert k[-1] == "d.b"
        assert v[-1] == "boo"

        del k[-1], v[-1]
        cf.reset_option("d.b")
        assert k[-1] == "d.b"

    def test_set_ContextManager(self):
        def eq(val):
            assert cf.get_option("a") == val

        cf.register_option("a", 0)
        eq(0)
        with cf.option_context("a", 15):
            eq(15)
            with cf.option_context("a", 25):
                eq(25)
            eq(15)
        eq(0)

        cf.set_option("a", 17)
        eq(17)

        # Test that option_context can be used as a decorator too (#34253).
        @cf.option_context("a", 123)
        def f():
            eq(123)

        f()

    def test_attribute_access(self):
        holder = []

        def f3(key):
            holder.append(True)

        cf.register_option("a", 0)
        cf.register_option("c", 0, cb=f3)
        options = cf.options

        assert options.a == 0
        with cf.option_context("a", 15):
            assert options.a == 15

        options.a = 500
        assert cf.get_option("a") == 500

        cf.reset_option("a")
        assert options.a == cf.get_option("a", 0)

        msg = "You can only set the value of existing options"
        with pytest.raises(OptionError, match=msg):
            options.b = 1
        with pytest.raises(OptionError, match=msg):
            options.display = 1

        # make sure callback kicks when using this form of setting
        options.c = 1
        assert len(holder) == 1

    def test_option_context_scope(self):
        # Ensure that creating a context does not affect the existing
        # environment as it is supposed to be used with the `with` statement.
        # See https://github.com/pandas-dev/pandas/issues/8514

        original_value = 60
        context_value = 10
        option_name = "a"

        cf.register_option(option_name, original_value)

        # Ensure creating contexts didn't affect the current context.
        ctx = cf.option_context(option_name, context_value)
        assert cf.get_option(option_name) == original_value

        # Ensure the correct value is available inside the context.
        with ctx:
            assert cf.get_option(option_name) == context_value

        # Ensure the current context is reset
        assert cf.get_option(option_name) == original_value

    def test_dictwrapper_getattr(self):
        options = cf.options
        # GH 19789
        with pytest.raises(OptionError, match="No such option"):
            options.bananas
        assert not hasattr(options, "bananas")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\groupby\aggregate\test_cython.py ===
"""
test cython .agg behavior
"""

import numpy as np
import pytest

from pandas.core.dtypes.common import (
    is_float_dtype,
    is_integer_dtype,
)

import pandas as pd
from pandas import (
    DataFrame,
    Index,
    NaT,
    Series,
    Timedelta,
    Timestamp,
    bdate_range,
)
import pandas._testing as tm
import pandas.core.common as com


@pytest.mark.parametrize(
    "op_name",
    [
        "count",
        "sum",
        "std",
        "var",
        "sem",
        "mean",
        pytest.param(
            "median",
            # ignore mean of empty slice
            # and all-NaN
            marks=[pytest.mark.filterwarnings("ignore::RuntimeWarning")],
        ),
        "prod",
        "min",
        "max",
    ],
)
def test_cythonized_aggers(op_name):
    data = {
        "A": [0, 0, 0, 0, 1, 1, 1, 1, 1, 1.0, np.nan, np.nan],
        "B": ["A", "B"] * 6,
        "C": np.random.default_rng(2).standard_normal(12),
    }
    df = DataFrame(data)
    df.loc[2:10:2, "C"] = np.nan

    op = lambda x: getattr(x, op_name)()

    # single column
    grouped = df.drop(["B"], axis=1).groupby("A")
    exp = {cat: op(group["C"]) for cat, group in grouped}
    exp = DataFrame({"C": exp})
    exp.index.name = "A"
    result = op(grouped)
    tm.assert_frame_equal(result, exp)

    # multiple columns
    grouped = df.groupby(["A", "B"])
    expd = {}
    for (cat1, cat2), group in grouped:
        expd.setdefault(cat1, {})[cat2] = op(group["C"])
    exp = DataFrame(expd).T.stack(future_stack=True)
    exp.index.names = ["A", "B"]
    exp.name = "C"

    result = op(grouped)["C"]
    if op_name in ["sum", "prod"]:
        tm.assert_series_equal(result, exp)


def test_cython_agg_boolean():
    frame = DataFrame(
        {
            "a": np.random.default_rng(2).integers(0, 5, 50),
            "b": np.random.default_rng(2).integers(0, 2, 50).astype("bool"),
        }
    )
    result = frame.groupby("a")["b"].mean()
    msg = "using SeriesGroupBy.mean"
    with tm.assert_produces_warning(FutureWarning, match=msg):
        # GH#53425
        expected = frame.groupby("a")["b"].agg(np.mean)

    tm.assert_series_equal(result, expected)


def test_cython_agg_nothing_to_agg():
    frame = DataFrame(
        {"a": np.random.default_rng(2).integers(0, 5, 50), "b": ["foo", "bar"] * 25}
    )

    msg = "Cannot use numeric_only=True with SeriesGroupBy.mean and non-numeric dtypes"
    with pytest.raises(TypeError, match=msg):
        frame.groupby("a")["b"].mean(numeric_only=True)

    frame = DataFrame(
        {"a": np.random.default_rng(2).integers(0, 5, 50), "b": ["foo", "bar"] * 25}
    )

    result = frame[["b"]].groupby(frame["a"]).mean(numeric_only=True)
    expected = DataFrame(
        [],
        index=frame["a"].sort_values().drop_duplicates(),
        columns=Index([], dtype="str"),
    )
    tm.assert_frame_equal(result, expected)


def test_cython_agg_nothing_to_agg_with_dates():
    frame = DataFrame(
        {
            "a": np.random.default_rng(2).integers(0, 5, 50),
            "b": ["foo", "bar"] * 25,
            "dates": pd.date_range("now", periods=50, freq="min"),
        }
    )
    msg = "Cannot use numeric_only=True with SeriesGroupBy.mean and non-numeric dtypes"
    with pytest.raises(TypeError, match=msg):
        frame.groupby("b").dates.mean(numeric_only=True)


def test_cython_agg_frame_columns():
    # #2113
    df = DataFrame({"x": [1, 2, 3], "y": [3, 4, 5]})

    msg = "DataFrame.groupby with axis=1 is deprecated"
    with tm.assert_produces_warning(FutureWarning, match=msg):
        df.groupby(level=0, axis="columns").mean()
    with tm.assert_produces_warning(FutureWarning, match=msg):
        df.groupby(level=0, axis="columns").mean()
    with tm.assert_produces_warning(FutureWarning, match=msg):
        df.groupby(level=0, axis="columns").mean()
    with tm.assert_produces_warning(FutureWarning, match=msg):
        df.groupby(level=0, axis="columns").mean()


def test_cython_agg_return_dict():
    # GH 16741
    df = DataFrame(
        {
            "A": ["foo", "bar", "foo", "bar", "foo", "bar", "foo", "foo"],
            "B": ["one", "one", "two", "three", "two", "two", "one", "three"],
            "C": np.random.default_rng(2).standard_normal(8),
            "D": np.random.default_rng(2).standard_normal(8),
        }
    )

    ts = df.groupby("A")["B"].agg(lambda x: x.value_counts().to_dict())
    expected = Series(
        [{"two": 1, "one": 1, "three": 1}, {"two": 2, "one": 2, "three": 1}],
        index=Index(["bar", "foo"], name="A"),
        name="B",
    )
    tm.assert_series_equal(ts, expected)


def test_cython_fail_agg():
    dr = bdate_range("1/1/2000", periods=50)
    ts = Series(["A", "B", "C", "D", "E"] * 10, dtype=object, index=dr)

    grouped = ts.groupby(lambda x: x.month)
    summed = grouped.sum()
    msg = "using SeriesGroupBy.sum"
    with tm.assert_produces_warning(FutureWarning, match=msg):
        # GH#53425
        expected = grouped.agg(np.sum).astype(object)
    tm.assert_series_equal(summed, expected)


@pytest.mark.parametrize(
    "op, targop",
    [
        ("mean", np.mean),
        ("median", np.median),
        ("var", np.var),
        ("sum", np.sum),
        ("prod", np.prod),
        ("min", np.min),
        ("max", np.max),
        ("first", lambda x: x.iloc[0]),
        ("last", lambda x: x.iloc[-1]),
    ],
)
def test__cython_agg_general(op, targop):
    df = DataFrame(np.random.default_rng(2).standard_normal(1000))
    labels = np.random.default_rng(2).integers(0, 50, size=1000).astype(float)

    result = df.groupby(labels)._cython_agg_general(op, alt=None, numeric_only=True)
    warn = FutureWarning if targop in com._cython_table else None
    msg = f"using DataFrameGroupBy.{op}"
    with tm.assert_produces_warning(warn, match=msg):
        # GH#53425
        expected = df.groupby(labels).agg(targop)
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "op, targop",
    [
        ("mean", np.mean),
        ("median", lambda x: np.median(x) if len(x) > 0 else np.nan),
        ("var", lambda x: np.var(x, ddof=1)),
        ("min", np.min),
        ("max", np.max),
    ],
)
def test_cython_agg_empty_buckets(op, targop, observed):
    df = DataFrame([11, 12, 13])
    grps = range(0, 55, 5)

    # calling _cython_agg_general directly, instead of via the user API
    # which sets different values for min_count, so do that here.
    g = df.groupby(pd.cut(df[0], grps), observed=observed)
    result = g._cython_agg_general(op, alt=None, numeric_only=True)

    g = df.groupby(pd.cut(df[0], grps), observed=observed)
    expected = g.agg(lambda x: targop(x))
    tm.assert_frame_equal(result, expected)


def test_cython_agg_empty_buckets_nanops(observed):
    # GH-18869 can't call nanops on empty groups, so hardcode expected
    # for these
    df = DataFrame([11, 12, 13], columns=["a"])
    grps = np.arange(0, 25, 5, dtype=int)
    # add / sum
    result = df.groupby(pd.cut(df["a"], grps), observed=observed)._cython_agg_general(
        "sum", alt=None, numeric_only=True
    )
    intervals = pd.interval_range(0, 20, freq=5)
    expected = DataFrame(
        {"a": [0, 0, 36, 0]},
        index=pd.CategoricalIndex(intervals, name="a", ordered=True),
    )
    if observed:
        expected = expected[expected.a != 0]

    tm.assert_frame_equal(result, expected)

    # prod
    result = df.groupby(pd.cut(df["a"], grps), observed=observed)._cython_agg_general(
        "prod", alt=None, numeric_only=True
    )
    expected = DataFrame(
        {"a": [1, 1, 1716, 1]},
        index=pd.CategoricalIndex(intervals, name="a", ordered=True),
    )
    if observed:
        expected = expected[expected.a != 1]

    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("op", ["first", "last", "max", "min"])
@pytest.mark.parametrize(
    "data", [Timestamp("2016-10-14 21:00:44.557"), Timedelta("17088 days 21:00:44.557")]
)
def test_cython_with_timestamp_and_nat(op, data):
    # https://github.com/pandas-dev/pandas/issues/19526
    df = DataFrame({"a": [0, 1], "b": [data, NaT]})
    index = Index([0, 1], name="a")

    # We will group by a and test the cython aggregations
    expected = DataFrame({"b": [data, NaT]}, index=index)

    result = df.groupby("a").aggregate(op)
    tm.assert_frame_equal(expected, result)


@pytest.mark.parametrize(
    "agg",
    [
        "min",
        "max",
        "count",
        "sum",
        "prod",
        "var",
        "mean",
        "median",
        "ohlc",
        "cumprod",
        "cumsum",
        "shift",
        "any",
        "all",
        "quantile",
        "first",
        "last",
        "rank",
        "cummin",
        "cummax",
    ],
)
def test_read_only_buffer_source_agg(agg):
    # https://github.com/pandas-dev/pandas/issues/36014
    df = DataFrame(
        {
            "sepal_length": [5.1, 4.9, 4.7, 4.6, 5.0],
            "species": ["setosa", "setosa", "setosa", "setosa", "setosa"],
        }
    )
    df._mgr.arrays[0].flags.writeable = False

    result = df.groupby(["species"]).agg({"sepal_length": agg})
    expected = df.copy().groupby(["species"]).agg({"sepal_length": agg})

    tm.assert_equal(result, expected)


@pytest.mark.parametrize(
    "op_name",
    [
        "count",
        "sum",
        "std",
        "var",
        "sem",
        "mean",
        "median",
        "prod",
        "min",
        "max",
    ],
)
def test_cython_agg_nullable_int(op_name):
    # ensure that the cython-based aggregations don't fail for nullable dtype
    # (eg https://github.com/pandas-dev/pandas/issues/37415)
    df = DataFrame(
        {
            "A": ["A", "B"] * 5,
            "B": pd.array([1, 2, 3, 4, 5, 6, 7, 8, 9, pd.NA], dtype="Int64"),
        }
    )
    result = getattr(df.groupby("A")["B"], op_name)()
    df2 = df.assign(B=df["B"].astype("float64"))
    expected = getattr(df2.groupby("A")["B"], op_name)()
    if op_name in ("mean", "median"):
        convert_integer = False
    else:
        convert_integer = True
    expected = expected.convert_dtypes(convert_integer=convert_integer)
    tm.assert_series_equal(result, expected)


@pytest.mark.parametrize("dtype", ["Int64", "Float64", "boolean"])
def test_count_masked_returns_masked_dtype(dtype):
    df = DataFrame(
        {
            "A": [1, 1],
            "B": pd.array([1, pd.NA], dtype=dtype),
            "C": pd.array([1, 1], dtype=dtype),
        }
    )
    result = df.groupby("A").count()
    expected = DataFrame(
        [[1, 2]], index=Index([1], name="A"), columns=["B", "C"], dtype="Int64"
    )
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("with_na", [True, False])
@pytest.mark.parametrize(
    "op_name, action",
    [
        # ("count", "always_int"),
        ("sum", "large_int"),
        # ("std", "always_float"),
        ("var", "always_float"),
        # ("sem", "always_float"),
        ("mean", "always_float"),
        ("median", "always_float"),
        ("prod", "large_int"),
        ("min", "preserve"),
        ("max", "preserve"),
        ("first", "preserve"),
        ("last", "preserve"),
    ],
)
@pytest.mark.parametrize(
    "data",
    [
        pd.array([1, 2, 3, 4], dtype="Int64"),
        pd.array([1, 2, 3, 4], dtype="Int8"),
        pd.array([0.1, 0.2, 0.3, 0.4], dtype="Float32"),
        pd.array([0.1, 0.2, 0.3, 0.4], dtype="Float64"),
        pd.array([True, True, False, False], dtype="boolean"),
    ],
)
def test_cython_agg_EA_known_dtypes(data, op_name, action, with_na):
    if with_na:
        data[3] = pd.NA

    df = DataFrame({"key": ["a", "a", "b", "b"], "col": data})
    grouped = df.groupby("key")

    if action == "always_int":
        # always Int64
        expected_dtype = pd.Int64Dtype()
    elif action == "large_int":
        # for any int/bool use Int64, for float preserve dtype
        if is_float_dtype(data.dtype):
            expected_dtype = data.dtype
        elif is_integer_dtype(data.dtype):
            # match the numpy dtype we'd get with the non-nullable analogue
            expected_dtype = data.dtype
        else:
            expected_dtype = pd.Int64Dtype()
    elif action == "always_float":
        # for any int/bool use Float64, for float preserve dtype
        if is_float_dtype(data.dtype):
            expected_dtype = data.dtype
        else:
            expected_dtype = pd.Float64Dtype()
    elif action == "preserve":
        expected_dtype = data.dtype

    result = getattr(grouped, op_name)()
    assert result["col"].dtype == expected_dtype

    result = grouped.aggregate(op_name)
    assert result["col"].dtype == expected_dtype

    result = getattr(grouped["col"], op_name)()
    assert result.dtype == expected_dtype

    result = grouped["col"].aggregate(op_name)
    assert result.dtype == expected_dtype

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\tseries\offsets\test_custom_business_month.py ===
"""
Tests for the following offsets:
- CustomBusinessMonthBase
- CustomBusinessMonthBegin
- CustomBusinessMonthEnd
"""
from __future__ import annotations

from datetime import (
    date,
    datetime,
    timedelta,
)

import numpy as np
import pytest

from pandas._libs.tslibs.offsets import (
    CBMonthBegin,
    CBMonthEnd,
    CDay,
)

import pandas._testing as tm
from pandas.tests.tseries.offsets.common import (
    assert_is_on_offset,
    assert_offset_equal,
)

from pandas.tseries import offsets


@pytest.fixture
def dt():
    return datetime(2008, 1, 1)


class TestCommonCBM:
    @pytest.mark.parametrize("offset2", [CBMonthBegin(2), CBMonthEnd(2)])
    def test_eq(self, offset2):
        assert offset2 == offset2

    @pytest.mark.parametrize("offset2", [CBMonthBegin(2), CBMonthEnd(2)])
    def test_hash(self, offset2):
        assert hash(offset2) == hash(offset2)

    @pytest.mark.parametrize("_offset", [CBMonthBegin, CBMonthEnd])
    def test_roundtrip_pickle(self, _offset):
        def _check_roundtrip(obj):
            unpickled = tm.round_trip_pickle(obj)
            assert unpickled == obj

        _check_roundtrip(_offset())
        _check_roundtrip(_offset(2))
        _check_roundtrip(_offset() * 2)

    @pytest.mark.parametrize("_offset", [CBMonthBegin, CBMonthEnd])
    def test_copy(self, _offset):
        # GH 17452
        off = _offset(weekmask="Mon Wed Fri")
        assert off == off.copy()


class TestCustomBusinessMonthBegin:
    @pytest.fixture
    def _offset(self):
        return CBMonthBegin

    @pytest.fixture
    def offset(self):
        return CBMonthBegin()

    @pytest.fixture
    def offset2(self):
        return CBMonthBegin(2)

    def test_different_normalize_equals(self, _offset):
        # GH#21404 changed __eq__ to return False when `normalize` does not match
        offset = _offset()
        offset2 = _offset(normalize=True)
        assert offset != offset2

    def test_repr(self, offset, offset2):
        assert repr(offset) == "<CustomBusinessMonthBegin>"
        assert repr(offset2) == "<2 * CustomBusinessMonthBegins>"

    def test_add_datetime(self, dt, offset2):
        assert offset2 + dt == datetime(2008, 3, 3)

    def testRollback1(self):
        assert CDay(10).rollback(datetime(2007, 12, 31)) == datetime(2007, 12, 31)

    def testRollback2(self, dt):
        assert CBMonthBegin(10).rollback(dt) == datetime(2008, 1, 1)

    def testRollforward1(self, dt):
        assert CBMonthBegin(10).rollforward(dt) == datetime(2008, 1, 1)

    def test_roll_date_object(self):
        offset = CBMonthBegin()

        dt = date(2012, 9, 15)

        result = offset.rollback(dt)
        assert result == datetime(2012, 9, 3)

        result = offset.rollforward(dt)
        assert result == datetime(2012, 10, 1)

        offset = offsets.Day()
        result = offset.rollback(dt)
        assert result == datetime(2012, 9, 15)

        result = offset.rollforward(dt)
        assert result == datetime(2012, 9, 15)

    on_offset_cases = [
        (CBMonthBegin(), datetime(2008, 1, 1), True),
        (CBMonthBegin(), datetime(2008, 1, 31), False),
    ]

    @pytest.mark.parametrize("case", on_offset_cases)
    def test_is_on_offset(self, case):
        offset, dt, expected = case
        assert_is_on_offset(offset, dt, expected)

    apply_cases = [
        (
            CBMonthBegin(),
            {
                datetime(2008, 1, 1): datetime(2008, 2, 1),
                datetime(2008, 2, 7): datetime(2008, 3, 3),
            },
        ),
        (
            2 * CBMonthBegin(),
            {
                datetime(2008, 1, 1): datetime(2008, 3, 3),
                datetime(2008, 2, 7): datetime(2008, 4, 1),
            },
        ),
        (
            -CBMonthBegin(),
            {
                datetime(2008, 1, 1): datetime(2007, 12, 3),
                datetime(2008, 2, 8): datetime(2008, 2, 1),
            },
        ),
        (
            -2 * CBMonthBegin(),
            {
                datetime(2008, 1, 1): datetime(2007, 11, 1),
                datetime(2008, 2, 9): datetime(2008, 1, 1),
            },
        ),
        (
            CBMonthBegin(0),
            {
                datetime(2008, 1, 1): datetime(2008, 1, 1),
                datetime(2008, 1, 7): datetime(2008, 2, 1),
            },
        ),
    ]

    @pytest.mark.parametrize("case", apply_cases)
    def test_apply(self, case):
        offset, cases = case
        for base, expected in cases.items():
            assert_offset_equal(offset, base, expected)

    def test_apply_large_n(self):
        dt = datetime(2012, 10, 23)

        result = dt + CBMonthBegin(10)
        assert result == datetime(2013, 8, 1)

        result = dt + CDay(100) - CDay(100)
        assert result == dt

        off = CBMonthBegin() * 6
        rs = datetime(2012, 1, 1) - off
        xp = datetime(2011, 7, 1)
        assert rs == xp

        st = datetime(2011, 12, 18)
        rs = st + off

        xp = datetime(2012, 6, 1)
        assert rs == xp

    def test_holidays(self):
        # Define a TradingDay offset
        holidays = ["2012-02-01", datetime(2012, 2, 2), np.datetime64("2012-03-01")]
        bm_offset = CBMonthBegin(holidays=holidays)
        dt = datetime(2012, 1, 1)

        assert dt + bm_offset == datetime(2012, 1, 2)
        assert dt + 2 * bm_offset == datetime(2012, 2, 3)

    @pytest.mark.parametrize(
        "case",
        [
            (
                CBMonthBegin(n=1, offset=timedelta(days=5)),
                {
                    datetime(2021, 3, 1): datetime(2021, 4, 1) + timedelta(days=5),
                    datetime(2021, 4, 17): datetime(2021, 5, 3) + timedelta(days=5),
                },
            ),
            (
                CBMonthBegin(n=2, offset=timedelta(days=40)),
                {
                    datetime(2021, 3, 10): datetime(2021, 5, 3) + timedelta(days=40),
                    datetime(2021, 4, 30): datetime(2021, 6, 1) + timedelta(days=40),
                },
            ),
            (
                CBMonthBegin(n=1, offset=timedelta(days=-5)),
                {
                    datetime(2021, 3, 1): datetime(2021, 4, 1) - timedelta(days=5),
                    datetime(2021, 4, 11): datetime(2021, 5, 3) - timedelta(days=5),
                },
            ),
            (
                -2 * CBMonthBegin(n=1, offset=timedelta(days=10)),
                {
                    datetime(2021, 3, 1): datetime(2021, 1, 1) + timedelta(days=10),
                    datetime(2021, 4, 3): datetime(2021, 3, 1) + timedelta(days=10),
                },
            ),
            (
                CBMonthBegin(n=0, offset=timedelta(days=1)),
                {
                    datetime(2021, 3, 2): datetime(2021, 4, 1) + timedelta(days=1),
                    datetime(2021, 4, 1): datetime(2021, 4, 1) + timedelta(days=1),
                },
            ),
            (
                CBMonthBegin(
                    n=1, holidays=["2021-04-01", "2021-04-02"], offset=timedelta(days=1)
                ),
                {
                    datetime(2021, 3, 2): datetime(2021, 4, 5) + timedelta(days=1),
                },
            ),
        ],
    )
    def test_apply_with_extra_offset(self, case):
        offset, cases = case
        for base, expected in cases.items():
            assert_offset_equal(offset, base, expected)


class TestCustomBusinessMonthEnd:
    @pytest.fixture
    def _offset(self):
        return CBMonthEnd

    @pytest.fixture
    def offset(self):
        return CBMonthEnd()

    @pytest.fixture
    def offset2(self):
        return CBMonthEnd(2)

    def test_different_normalize_equals(self, _offset):
        # GH#21404 changed __eq__ to return False when `normalize` does not match
        offset = _offset()
        offset2 = _offset(normalize=True)
        assert offset != offset2

    def test_repr(self, offset, offset2):
        assert repr(offset) == "<CustomBusinessMonthEnd>"
        assert repr(offset2) == "<2 * CustomBusinessMonthEnds>"

    def test_add_datetime(self, dt, offset2):
        assert offset2 + dt == datetime(2008, 2, 29)

    def testRollback1(self):
        assert CDay(10).rollback(datetime(2007, 12, 31)) == datetime(2007, 12, 31)

    def testRollback2(self, dt):
        assert CBMonthEnd(10).rollback(dt) == datetime(2007, 12, 31)

    def testRollforward1(self, dt):
        assert CBMonthEnd(10).rollforward(dt) == datetime(2008, 1, 31)

    def test_roll_date_object(self):
        offset = CBMonthEnd()

        dt = date(2012, 9, 15)

        result = offset.rollback(dt)
        assert result == datetime(2012, 8, 31)

        result = offset.rollforward(dt)
        assert result == datetime(2012, 9, 28)

        offset = offsets.Day()
        result = offset.rollback(dt)
        assert result == datetime(2012, 9, 15)

        result = offset.rollforward(dt)
        assert result == datetime(2012, 9, 15)

    on_offset_cases = [
        (CBMonthEnd(), datetime(2008, 1, 31), True),
        (CBMonthEnd(), datetime(2008, 1, 1), False),
    ]

    @pytest.mark.parametrize("case", on_offset_cases)
    def test_is_on_offset(self, case):
        offset, dt, expected = case
        assert_is_on_offset(offset, dt, expected)

    apply_cases = [
        (
            CBMonthEnd(),
            {
                datetime(2008, 1, 1): datetime(2008, 1, 31),
                datetime(2008, 2, 7): datetime(2008, 2, 29),
            },
        ),
        (
            2 * CBMonthEnd(),
            {
                datetime(2008, 1, 1): datetime(2008, 2, 29),
                datetime(2008, 2, 7): datetime(2008, 3, 31),
            },
        ),
        (
            -CBMonthEnd(),
            {
                datetime(2008, 1, 1): datetime(2007, 12, 31),
                datetime(2008, 2, 8): datetime(2008, 1, 31),
            },
        ),
        (
            -2 * CBMonthEnd(),
            {
                datetime(2008, 1, 1): datetime(2007, 11, 30),
                datetime(2008, 2, 9): datetime(2007, 12, 31),
            },
        ),
        (
            CBMonthEnd(0),
            {
                datetime(2008, 1, 1): datetime(2008, 1, 31),
                datetime(2008, 2, 7): datetime(2008, 2, 29),
            },
        ),
    ]

    @pytest.mark.parametrize("case", apply_cases)
    def test_apply(self, case):
        offset, cases = case
        for base, expected in cases.items():
            assert_offset_equal(offset, base, expected)

    def test_apply_large_n(self):
        dt = datetime(2012, 10, 23)

        result = dt + CBMonthEnd(10)
        assert result == datetime(2013, 7, 31)

        result = dt + CDay(100) - CDay(100)
        assert result == dt

        off = CBMonthEnd() * 6
        rs = datetime(2012, 1, 1) - off
        xp = datetime(2011, 7, 29)
        assert rs == xp

        st = datetime(2011, 12, 18)
        rs = st + off
        xp = datetime(2012, 5, 31)
        assert rs == xp

    def test_holidays(self):
        # Define a TradingDay offset
        holidays = ["2012-01-31", datetime(2012, 2, 28), np.datetime64("2012-02-29")]
        bm_offset = CBMonthEnd(holidays=holidays)
        dt = datetime(2012, 1, 1)
        assert dt + bm_offset == datetime(2012, 1, 30)
        assert dt + 2 * bm_offset == datetime(2012, 2, 27)

    @pytest.mark.parametrize(
        "case",
        [
            (
                CBMonthEnd(n=1, offset=timedelta(days=5)),
                {
                    datetime(2021, 3, 1): datetime(2021, 3, 31) + timedelta(days=5),
                    datetime(2021, 4, 17): datetime(2021, 4, 30) + timedelta(days=5),
                },
            ),
            (
                CBMonthEnd(n=2, offset=timedelta(days=40)),
                {
                    datetime(2021, 3, 10): datetime(2021, 4, 30) + timedelta(days=40),
                    datetime(2021, 4, 30): datetime(2021, 6, 30) + timedelta(days=40),
                },
            ),
            (
                CBMonthEnd(n=1, offset=timedelta(days=-5)),
                {
                    datetime(2021, 3, 1): datetime(2021, 3, 31) - timedelta(days=5),
                    datetime(2021, 4, 11): datetime(2021, 4, 30) - timedelta(days=5),
                },
            ),
            (
                -2 * CBMonthEnd(n=1, offset=timedelta(days=10)),
                {
                    datetime(2021, 3, 1): datetime(2021, 1, 29) + timedelta(days=10),
                    datetime(2021, 4, 3): datetime(2021, 2, 26) + timedelta(days=10),
                },
            ),
            (
                CBMonthEnd(n=0, offset=timedelta(days=1)),
                {
                    datetime(2021, 3, 2): datetime(2021, 3, 31) + timedelta(days=1),
                    datetime(2021, 4, 1): datetime(2021, 4, 30) + timedelta(days=1),
                },
            ),
            (
                CBMonthEnd(n=1, holidays=["2021-03-31"], offset=timedelta(days=1)),
                {
                    datetime(2021, 3, 2): datetime(2021, 3, 30) + timedelta(days=1),
                },
            ),
        ],
    )
    def test_apply_with_extra_offset(self, case):
        offset, cases = case
        for base, expected in cases.items():
            assert_offset_equal(offset, base, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\algebras\tests\test_quaternion.py ===
from sympy.testing.pytest import slow
from sympy.core.function import diff
from sympy.core.function import expand
from sympy.core.numbers import (E, I, Rational, pi)
from sympy.core.singleton import S
from sympy.core.symbol import (Symbol, symbols)
from sympy.functions.elementary.complexes import (Abs, conjugate, im, re, sign)
from sympy.functions.elementary.exponential import log
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.trigonometric import (acos, asin, cos, sin, atan2, atan)
from sympy.integrals.integrals import integrate
from sympy.matrices.dense import Matrix
from sympy.simplify import simplify
from sympy.simplify.trigsimp import trigsimp
from sympy.algebras.quaternion import Quaternion
from sympy.testing.pytest import raises
import math
from itertools import permutations, product

w, x, y, z = symbols('w:z')
phi = symbols('phi')

def test_quaternion_construction():
    q = Quaternion(w, x, y, z)
    assert q + q == Quaternion(2*w, 2*x, 2*y, 2*z)

    q2 = Quaternion.from_axis_angle((sqrt(3)/3, sqrt(3)/3, sqrt(3)/3),
                                    pi*Rational(2, 3))
    assert q2 == Quaternion(S.Half, S.Half,
                            S.Half, S.Half)

    M = Matrix([[cos(phi), -sin(phi), 0], [sin(phi), cos(phi), 0], [0, 0, 1]])
    q3 = trigsimp(Quaternion.from_rotation_matrix(M))
    assert q3 == Quaternion(
        sqrt(2)*sqrt(cos(phi) + 1)/2, 0, 0, sqrt(2 - 2*cos(phi))*sign(sin(phi))/2)

    nc = Symbol('nc', commutative=False)
    raises(ValueError, lambda: Quaternion(w, x, nc, z))


def test_quaternion_construction_norm():
    q1 = Quaternion(*symbols('a:d'))

    q2 = Quaternion(w, x, y, z)
    assert expand((q1*q2).norm()**2 - (q1.norm()**2 * q2.norm()**2)) == 0

    q3 = Quaternion(w, x, y, z, norm=1)
    assert (q1 * q3).norm() == q1.norm()


def test_issue_25254():
    # calculating the inverse cached the norm which caused problems
    # when multiplying
    p = Quaternion(1, 0, 0, 0)
    q = Quaternion.from_axis_angle((1, 1, 1), 3 * math.pi/4)
    qi = q.inverse()  # this operation cached the norm
    test = q * p * qi
    assert ((test - p).norm() < 1E-10)


def test_to_and_from_Matrix():
    q = Quaternion(w, x, y, z)
    q_full = Quaternion.from_Matrix(q.to_Matrix())
    q_vect = Quaternion.from_Matrix(q.to_Matrix(True))
    assert (q - q_full).is_zero_quaternion()
    assert (q.vector_part() - q_vect).is_zero_quaternion()


def test_product_matrices():
    q1 = Quaternion(w, x, y, z)
    q2 = Quaternion(*(symbols("a:d")))
    assert (q1 * q2).to_Matrix() == q1.product_matrix_left * q2.to_Matrix()
    assert (q1 * q2).to_Matrix() == q2.product_matrix_right * q1.to_Matrix()

    R1 = (q1.product_matrix_left * q1.product_matrix_right.T)[1:, 1:]
    R2 = simplify(q1.to_rotation_matrix()*q1.norm()**2)
    assert R1 == R2


def test_quaternion_axis_angle():

    test_data = [ # axis, angle, expected_quaternion
        ((1, 0, 0), 0, (1, 0, 0, 0)),
        ((1, 0, 0), pi/2, (sqrt(2)/2, sqrt(2)/2, 0, 0)),
        ((0, 1, 0), pi/2, (sqrt(2)/2, 0, sqrt(2)/2, 0)),
        ((0, 0, 1), pi/2, (sqrt(2)/2, 0, 0, sqrt(2)/2)),
        ((1, 0, 0), pi, (0, 1, 0, 0)),
        ((0, 1, 0), pi, (0, 0, 1, 0)),
        ((0, 0, 1), pi, (0, 0, 0, 1)),
        ((1, 1, 1), pi, (0, 1/sqrt(3),1/sqrt(3),1/sqrt(3))),
        ((sqrt(3)/3, sqrt(3)/3, sqrt(3)/3), pi*2/3, (S.Half, S.Half, S.Half, S.Half))
    ]

    for axis, angle, expected in test_data:
        assert Quaternion.from_axis_angle(axis, angle) == Quaternion(*expected)


def test_quaternion_axis_angle_simplification():
    result = Quaternion.from_axis_angle((1, 2, 3), asin(4))
    assert result.a == cos(asin(4)/2)
    assert result.b == sqrt(14)*sin(asin(4)/2)/14
    assert result.c == sqrt(14)*sin(asin(4)/2)/7
    assert result.d == 3*sqrt(14)*sin(asin(4)/2)/14

def test_quaternion_complex_real_addition():
    a = symbols("a", complex=True)
    b = symbols("b", real=True)
    # This symbol is not complex:
    c = symbols("c", commutative=False)

    q = Quaternion(w, x, y, z)
    assert a + q == Quaternion(w + re(a), x + im(a), y, z)
    assert 1 + q == Quaternion(1 + w, x, y, z)
    assert I + q == Quaternion(w, 1 + x, y, z)
    assert b + q == Quaternion(w + b, x, y, z)
    raises(ValueError, lambda: c + q)
    raises(ValueError, lambda: q * c)
    raises(ValueError, lambda: c * q)

    assert -q == Quaternion(-w, -x, -y, -z)

    q1 = Quaternion(3 + 4*I, 2 + 5*I, 0, 7 + 8*I, real_field = False)
    q2 = Quaternion(1, 4, 7, 8)

    assert q1 + (2 + 3*I) == Quaternion(5 + 7*I, 2 + 5*I, 0, 7 + 8*I)
    assert q2 + (2 + 3*I) == Quaternion(3, 7, 7, 8)
    assert q1 * (2 + 3*I) == \
    Quaternion((2 + 3*I)*(3 + 4*I), (2 + 3*I)*(2 + 5*I), 0, (2 + 3*I)*(7 + 8*I))
    assert q2 * (2 + 3*I) == Quaternion(-10, 11, 38, -5)

    q1 = Quaternion(1, 2, 3, 4)
    q0 = Quaternion(0, 0, 0, 0)
    assert q1 + q0 == q1
    assert q1 - q0 == q1
    assert q1 - q1 == q0


def test_quaternion_subs():
    q = Quaternion.from_axis_angle((0, 0, 1), phi)
    assert q.subs(phi, 0) == Quaternion(1, 0, 0, 0)


def test_quaternion_evalf():
    assert (Quaternion(sqrt(2), 0, 0, sqrt(3)).evalf() ==
            Quaternion(sqrt(2).evalf(), 0, 0, sqrt(3).evalf()))
    assert (Quaternion(1/sqrt(2), 0, 0, 1/sqrt(2)).evalf() ==
            Quaternion((1/sqrt(2)).evalf(), 0, 0, (1/sqrt(2)).evalf()))


def test_quaternion_functions():
    q = Quaternion(w, x, y, z)
    q1 = Quaternion(1, 2, 3, 4)
    q0 = Quaternion(0, 0, 0, 0)

    assert conjugate(q) == Quaternion(w, -x, -y, -z)
    assert q.norm() == sqrt(w**2 + x**2 + y**2 + z**2)
    assert q.normalize() == Quaternion(w, x, y, z) / sqrt(w**2 + x**2 + y**2 + z**2)
    assert q.inverse() == Quaternion(w, -x, -y, -z) / (w**2 + x**2 + y**2 + z**2)
    assert q.inverse() == q.pow(-1)
    raises(ValueError, lambda: q0.inverse())
    assert q.pow(2) == Quaternion(w**2 - x**2 - y**2 - z**2, 2*w*x, 2*w*y, 2*w*z)
    assert q**(2) == Quaternion(w**2 - x**2 - y**2 - z**2, 2*w*x, 2*w*y, 2*w*z)
    assert q1.pow(-2) == Quaternion(
        Rational(-7, 225), Rational(-1, 225), Rational(-1, 150), Rational(-2, 225))
    assert q1**(-2) == Quaternion(
        Rational(-7, 225), Rational(-1, 225), Rational(-1, 150), Rational(-2, 225))
    assert q1.pow(-0.5) == NotImplemented
    raises(TypeError, lambda: q1**(-0.5))

    assert q1.exp() == \
    Quaternion(E * cos(sqrt(29)),
               2 * sqrt(29) * E * sin(sqrt(29)) / 29,
               3 * sqrt(29) * E * sin(sqrt(29)) / 29,
               4 * sqrt(29) * E * sin(sqrt(29)) / 29)
    assert q1.log() == \
    Quaternion(log(sqrt(30)),
               2 * sqrt(29) * acos(sqrt(30)/30) / 29,
               3 * sqrt(29) * acos(sqrt(30)/30) / 29,
               4 * sqrt(29) * acos(sqrt(30)/30) / 29)

    assert q1.pow_cos_sin(2) == \
    Quaternion(30 * cos(2 * acos(sqrt(30)/30)),
               60 * sqrt(29) * sin(2 * acos(sqrt(30)/30)) / 29,
               90 * sqrt(29) * sin(2 * acos(sqrt(30)/30)) / 29,
               120 * sqrt(29) * sin(2 * acos(sqrt(30)/30)) / 29)

    assert diff(Quaternion(x, x, x, x), x) == Quaternion(1, 1, 1, 1)

    assert integrate(Quaternion(x, x, x, x), x) == \
    Quaternion(x**2 / 2, x**2 / 2, x**2 / 2, x**2 / 2)

    assert Quaternion(1, x, x**2, x**3).integrate(x) == \
    Quaternion(x, x**2/2, x**3/3, x**4/4)

    assert Quaternion(sin(x), cos(x), sin(2*x), cos(2*x)).integrate(x) == \
    Quaternion(-cos(x), sin(x), -cos(2*x)/2, sin(2*x)/2)

    assert Quaternion(x**2, y**2, z**2, x*y*z).integrate(x, y) == \
    Quaternion(x**3*y/3, x*y**3/3, x*y*z**2, x**2*y**2*z/4)

    assert Quaternion.rotate_point((1, 1, 1), q1) == (S.One / 5, 1, S(7) / 5)
    n = Symbol('n')
    raises(TypeError, lambda: q1**n)
    n = Symbol('n', integer=True)
    raises(TypeError, lambda: q1**n)

    assert Quaternion(22, 23, 55, 8).scalar_part() == 22
    assert Quaternion(w, x, y, z).scalar_part() == w

    assert Quaternion(22, 23, 55, 8).vector_part() == Quaternion(0, 23, 55, 8)
    assert Quaternion(w, x, y, z).vector_part() == Quaternion(0, x, y, z)

    assert q1.axis() == Quaternion(0, 2*sqrt(29)/29, 3*sqrt(29)/29, 4*sqrt(29)/29)
    assert q1.axis().pow(2) == Quaternion(-1, 0, 0, 0)
    assert q0.axis().scalar_part() == 0
    assert (q.axis() == Quaternion(0,
                                   x/sqrt(x**2 + y**2 + z**2),
                                   y/sqrt(x**2 + y**2 + z**2),
                                   z/sqrt(x**2 + y**2 + z**2)))

    assert q0.is_pure() is True
    assert q1.is_pure() is False
    assert Quaternion(0, 0, 0, 3).is_pure() is True
    assert Quaternion(0, 2, 10, 3).is_pure() is True
    assert Quaternion(w, 2, 10, 3).is_pure() is None

    assert q1.angle() == 2*atan(sqrt(29))
    assert q.angle() == 2*atan2(sqrt(x**2 + y**2 + z**2), w)

    assert Quaternion.arc_coplanar(q1, Quaternion(2, 4, 6, 8)) is True
    assert Quaternion.arc_coplanar(q1, Quaternion(1, -2, -3, -4)) is True
    assert Quaternion.arc_coplanar(q1, Quaternion(1, 8, 12, 16)) is True
    assert Quaternion.arc_coplanar(q1, Quaternion(1, 2, 3, 4)) is True
    assert Quaternion.arc_coplanar(q1, Quaternion(w, 4, 6, 8)) is True
    assert Quaternion.arc_coplanar(q1, Quaternion(2, 7, 4, 1)) is False
    assert Quaternion.arc_coplanar(q1, Quaternion(w, x, y, z)) is None
    raises(ValueError, lambda: Quaternion.arc_coplanar(q1, q0))

    assert Quaternion.vector_coplanar(
        Quaternion(0, 8, 12, 16),
        Quaternion(0, 4, 6, 8),
        Quaternion(0, 2, 3, 4)) is True
    assert Quaternion.vector_coplanar(
        Quaternion(0, 0, 0, 0), Quaternion(0, 4, 6, 8), Quaternion(0, 2, 3, 4)) is True
    assert Quaternion.vector_coplanar(
        Quaternion(0, 8, 2, 6), Quaternion(0, 1, 6, 6), Quaternion(0, 0, 3, 4)) is False
    assert Quaternion.vector_coplanar(
        Quaternion(0, 1, 3, 4),
        Quaternion(0, 4, w, 6),
        Quaternion(0, 6, 8, 1)) is None
    raises(ValueError, lambda:
        Quaternion.vector_coplanar(q0, Quaternion(0, 4, 6, 8), q1))

    assert Quaternion(0, 1, 2, 3).parallel(Quaternion(0, 2, 4, 6)) is True
    assert Quaternion(0, 1, 2, 3).parallel(Quaternion(0, 2, 2, 6)) is False
    assert Quaternion(0, 1, 2, 3).parallel(Quaternion(w, x, y, 6)) is None
    raises(ValueError, lambda: q0.parallel(q1))

    assert Quaternion(0, 1, 2, 3).orthogonal(Quaternion(0, -2, 1, 0)) is True
    assert Quaternion(0, 2, 4, 7).orthogonal(Quaternion(0, 2, 2, 6)) is False
    assert Quaternion(0, 2, 4, 7).orthogonal(Quaternion(w, x, y, 6)) is None
    raises(ValueError, lambda: q0.orthogonal(q1))

    assert q1.index_vector() == Quaternion(
        0, 2*sqrt(870)/29,
        3*sqrt(870)/29,
        4*sqrt(870)/29)
    assert Quaternion(0, 3, 9, 4).index_vector() == Quaternion(0, 3, 9, 4)

    assert Quaternion(4, 3, 9, 4).mensor() == log(sqrt(122))
    assert Quaternion(3, 3, 0, 2).mensor() == log(sqrt(22))

    assert q0.is_zero_quaternion() is True
    assert q1.is_zero_quaternion() is False
    assert Quaternion(w, 0, 0, 0).is_zero_quaternion() is None

def test_quaternion_conversions():
    q1 = Quaternion(1, 2, 3, 4)

    assert q1.to_axis_angle() == ((2 * sqrt(29)/29,
                                   3 * sqrt(29)/29,
                                   4 * sqrt(29)/29),
                                   2 * acos(sqrt(30)/30))

    assert (q1.to_rotation_matrix() ==
            Matrix([[Rational(-2, 3), Rational(2, 15), Rational(11, 15)],
                    [Rational(2, 3), Rational(-1, 3), Rational(2, 3)],
                    [Rational(1, 3), Rational(14, 15), Rational(2, 15)]]))

    assert (q1.to_rotation_matrix((1, 1, 1)) ==
            Matrix([
                [Rational(-2, 3), Rational(2, 15), Rational(11, 15), Rational(4, 5)],
                [Rational(2, 3), Rational(-1, 3), Rational(2, 3), S.Zero],
                [Rational(1, 3), Rational(14, 15), Rational(2, 15), Rational(-2, 5)],
                [S.Zero, S.Zero, S.Zero, S.One]]))

    theta = symbols("theta", real=True)
    q2 = Quaternion(cos(theta/2), 0, 0, sin(theta/2))

    assert trigsimp(q2.to_rotation_matrix()) == Matrix([
                                               [cos(theta), -sin(theta), 0],
                                               [sin(theta),  cos(theta), 0],
                                               [0,           0,          1]])

    assert q2.to_axis_angle() == ((0, 0, sin(theta/2)/Abs(sin(theta/2))),
                                   2*acos(cos(theta/2)))

    assert trigsimp(q2.to_rotation_matrix((1, 1, 1))) == Matrix([
               [cos(theta), -sin(theta), 0, sin(theta) - cos(theta) + 1],
               [sin(theta),  cos(theta), 0, -sin(theta) - cos(theta) + 1],
               [0,           0,          1,  0],
               [0,           0,          0,  1]])


def test_rotation_matrix_homogeneous():
    q = Quaternion(w, x, y, z)
    R1 = q.to_rotation_matrix(homogeneous=True) * q.norm()**2
    R2 = simplify(q.to_rotation_matrix(homogeneous=False) * q.norm()**2)
    assert R1 == R2


def test_quaternion_rotation_iss1593():
    """
    There was a sign mistake in the definition,
    of the rotation matrix. This tests that particular sign mistake.
    See issue 1593 for reference.
    See wikipedia
    https://en.wikipedia.org/wiki/Quaternions_and_spatial_rotation#Quaternion-derived_rotation_matrix
    for the correct definition
    """
    q = Quaternion(cos(phi/2), sin(phi/2), 0, 0)
    assert(trigsimp(q.to_rotation_matrix()) == Matrix([
                [1,        0,         0],
                [0, cos(phi), -sin(phi)],
                [0, sin(phi),  cos(phi)]]))


def test_quaternion_multiplication():
    q1 = Quaternion(3 + 4*I, 2 + 5*I, 0, 7 + 8*I, real_field = False)
    q2 = Quaternion(1, 2, 3, 5)
    q3 = Quaternion(1, 1, 1, y)

    assert Quaternion._generic_mul(S(4), S.One) == 4
    assert (Quaternion._generic_mul(S(4), q1) ==
            Quaternion(12 + 16*I, 8 + 20*I, 0, 28 + 32*I))
    assert q2.mul(2) == Quaternion(2, 4, 6, 10)
    assert q2.mul(q3) == Quaternion(-5*y - 4, 3*y - 2, 9 - 2*y, y + 4)
    assert q2.mul(q3) == q2*q3

    z = symbols('z', complex=True)
    z_quat = Quaternion(re(z), im(z), 0, 0)
    q = Quaternion(*symbols('q:4', real=True))

    assert z * q == z_quat * q
    assert q * z == q * z_quat


def test_issue_16318():
    #for rtruediv
    q0 = Quaternion(0, 0, 0, 0)
    raises(ValueError, lambda: 1/q0)
    #for rotate_point
    q = Quaternion(1, 2, 3, 4)
    (axis, angle) = q.to_axis_angle()
    assert Quaternion.rotate_point((1, 1, 1), (axis, angle)) == (S.One / 5, 1, S(7) / 5)
    #test for to_axis_angle
    q = Quaternion(-1, 1, 1, 1)
    axis = (-sqrt(3)/3, -sqrt(3)/3, -sqrt(3)/3)
    angle = 2*pi/3
    assert (axis, angle) == q.to_axis_angle()


@slow
def test_to_euler():
    q = Quaternion(w, x, y, z)
    q_normalized = q.normalize()

    seqs = ['zxy', 'zyx', 'zyz', 'zxz']
    seqs += [seq.upper() for seq in seqs]

    for seq in seqs:
        euler_from_q = q.to_euler(seq)
        q_back = simplify(Quaternion.from_euler(euler_from_q, seq))
        assert q_back == q_normalized


def test_to_euler_iss24504():
    """
    There was a mistake in the degenerate case testing
    See issue 24504 for reference.
    """
    q = Quaternion.from_euler((phi, 0, 0), 'zyz')
    assert trigsimp(q.to_euler('zyz'), inverse=True) == (phi, 0, 0)


def test_to_euler_numerical_singilarities():

    def test_one_case(angles, seq):
        q = Quaternion.from_euler(angles, seq)
        assert q.to_euler(seq) == angles

    # symmetric
    test_one_case((pi/2,  0, 0), 'zyz')
    test_one_case((pi/2,  0, 0), 'ZYZ')
    test_one_case((pi/2,  pi, 0), 'zyz')
    test_one_case((pi/2,  pi, 0), 'ZYZ')

    # asymmetric
    test_one_case((pi/2,  pi/2, 0), 'zyx')
    test_one_case((pi/2,  -pi/2, 0), 'zyx')
    test_one_case((pi/2,  pi/2, 0), 'ZYX')
    test_one_case((pi/2,  -pi/2, 0), 'ZYX')


@slow
def test_to_euler_options():
    def test_one_case(q):
        angles1 = Matrix(q.to_euler(seq, True, True))
        angles2 = Matrix(q.to_euler(seq, False, False))
        angle_errors = simplify(angles1-angles2).evalf()
        for angle_error in angle_errors:
            # forcing angles to set {-pi, pi}
            angle_error = (angle_error + pi) % (2 * pi) - pi
            assert angle_error < 10e-7

    for xyz in ('xyz', 'XYZ'):
        for seq_tuple in permutations(xyz):
            for symmetric in (True, False):
                if symmetric:
                    seq = ''.join([seq_tuple[0], seq_tuple[1], seq_tuple[0]])
                else:
                    seq = ''.join(seq_tuple)

                for elements in product([-1, 0, 1], repeat=4):
                    q = Quaternion(*elements)
                    if not q.is_zero_quaternion():
                        test_one_case(q)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\stats\tests\test_joint_rv.py ===
from sympy.concrete.products import Product
from sympy.concrete.summations import Sum
from sympy.core.numbers import (Rational, oo, pi)
from sympy.core.relational import Eq
from sympy.core.singleton import S
from sympy.core.symbol import symbols
from sympy.functions.combinatorial.factorials import (RisingFactorial, factorial)
from sympy.functions.elementary.complexes import polar_lift
from sympy.functions.elementary.exponential import exp
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.piecewise import Piecewise
from sympy.functions.special.bessel import besselk
from sympy.functions.special.gamma_functions import gamma
from sympy.matrices.dense import eye
from sympy.matrices.expressions.determinant import Determinant
from sympy.sets.fancysets import Range
from sympy.sets.sets import (Interval, ProductSet)
from sympy.simplify.simplify import simplify
from sympy.tensor.indexed import (Indexed, IndexedBase)
from sympy.core.numbers import comp
from sympy.integrals.integrals import integrate
from sympy.matrices import Matrix, MatrixSymbol
from sympy.matrices.expressions.matexpr import MatrixElement
from sympy.stats import density, median, marginal_distribution, Normal, Laplace, E, sample
from sympy.stats.joint_rv_types import (JointRV, MultivariateNormalDistribution,
                JointDistributionHandmade, MultivariateT, NormalGamma,
                GeneralizedMultivariateLogGammaOmega as GMVLGO, MultivariateBeta,
                GeneralizedMultivariateLogGamma as GMVLG, MultivariateEwens,
                Multinomial, NegativeMultinomial, MultivariateNormal,
                MultivariateLaplace)
from sympy.testing.pytest import raises, XFAIL, skip, slow
from sympy.external import import_module

from sympy.abc import x, y



def test_Normal():
    m = Normal('A', [1, 2], [[1, 0], [0, 1]])
    A = MultivariateNormal('A', [1, 2], [[1, 0], [0, 1]])
    assert m == A
    assert density(m)(1, 2) == 1/(2*pi)
    assert m.pspace.distribution.set == ProductSet(S.Reals, S.Reals)
    raises (ValueError, lambda:m[2])
    n = Normal('B', [1, 2, 3], [[1, 0, 0], [0, 1, 0], [0, 0, 1]])
    p = Normal('C',  Matrix([1, 2]), Matrix([[1, 0], [0, 1]]))
    assert density(m)(x, y) == density(p)(x, y)
    assert marginal_distribution(n, 0, 1)(1, 2) == 1/(2*pi)
    raises(ValueError, lambda: marginal_distribution(m))
    assert integrate(density(m)(x, y), (x, -oo, oo), (y, -oo, oo)).evalf() == 1.0
    N = Normal('N', [1, 2], [[x, 0], [0, y]])
    assert density(N)(0, 0) == exp(-((4*x + y)/(2*x*y)))/(2*pi*sqrt(x*y))

    raises (ValueError, lambda: Normal('M', [1, 2], [[1, 1], [1, -1]]))
    # symbolic
    n = symbols('n', integer=True, positive=True)
    mu = MatrixSymbol('mu', n, 1)
    sigma = MatrixSymbol('sigma', n, n)
    X = Normal('X', mu, sigma)
    assert density(X) == MultivariateNormalDistribution(mu, sigma)
    raises (NotImplementedError, lambda: median(m))
    # Below tests should work after issue #17267 is resolved
    # assert E(X) == mu
    # assert variance(X) == sigma

    # test symbolic multivariate normal densities
    n = 3

    Sg = MatrixSymbol('Sg', n, n)
    mu = MatrixSymbol('mu', n, 1)
    obs = MatrixSymbol('obs', n, 1)

    X = MultivariateNormal('X', mu, Sg)
    density_X = density(X)

    eval_a = density_X(obs).subs({Sg: eye(3),
        mu: Matrix([0, 0, 0]), obs: Matrix([0, 0, 0])}).doit()
    eval_b = density_X(0, 0, 0).subs({Sg: eye(3), mu: Matrix([0, 0, 0])}).doit()

    assert eval_a == sqrt(2)/(4*pi**Rational(3/2))
    assert eval_b == sqrt(2)/(4*pi**Rational(3/2))

    n = symbols('n', integer=True, positive=True)

    Sg = MatrixSymbol('Sg', n, n)
    mu = MatrixSymbol('mu', n, 1)
    obs = MatrixSymbol('obs', n, 1)

    X = MultivariateNormal('X', mu, Sg)
    density_X_at_obs = density(X)(obs)

    expected_density = MatrixElement(
        exp((S(1)/2) * (mu.T - obs.T) * Sg**(-1) * (-mu + obs)) / \
        sqrt((2*pi)**n * Determinant(Sg)), 0, 0)

    assert density_X_at_obs == expected_density


def test_MultivariateTDist():
    t1 = MultivariateT('T', [0, 0], [[1, 0], [0, 1]], 2)
    assert(density(t1))(1, 1) == 1/(8*pi)
    assert t1.pspace.distribution.set == ProductSet(S.Reals, S.Reals)
    assert integrate(density(t1)(x, y), (x, -oo, oo), \
        (y, -oo, oo)).evalf() == 1.0
    raises(ValueError, lambda: MultivariateT('T', [1, 2], [[1, 1], [1, -1]], 1))
    t2 = MultivariateT('t2', [1, 2], [[x, 0], [0, y]], 1)
    assert density(t2)(1, 2) == 1/(2*pi*sqrt(x*y))


def test_multivariate_laplace():
    raises(ValueError, lambda: Laplace('T', [1, 2], [[1, 2], [2, 1]]))
    L = Laplace('L', [1, 0], [[1, 0], [0, 1]])
    L2 = MultivariateLaplace('L2', [1, 0], [[1, 0], [0, 1]])
    assert density(L)(2, 3) == exp(2)*besselk(0, sqrt(39))/pi
    L1 = Laplace('L1', [1, 2], [[x, 0], [0, y]])
    assert density(L1)(0, 1) == \
        exp(2/y)*besselk(0, sqrt((2 + 4/y + 1/x)/y))/(pi*sqrt(x*y))
    assert L.pspace.distribution.set == ProductSet(S.Reals, S.Reals)
    assert L.pspace.distribution == L2.pspace.distribution


def test_NormalGamma():
    ng = NormalGamma('G', 1, 2, 3, 4)
    assert density(ng)(1, 1) == 32*exp(-4)/sqrt(pi)
    assert ng.pspace.distribution.set == ProductSet(S.Reals, Interval(0, oo))
    raises(ValueError, lambda:NormalGamma('G', 1, 2, 3, -1))
    assert marginal_distribution(ng, 0)(1) == \
        3*sqrt(10)*gamma(Rational(7, 4))/(10*sqrt(pi)*gamma(Rational(5, 4)))
    assert marginal_distribution(ng, y)(1) == exp(Rational(-1, 4))/128
    assert marginal_distribution(ng,[0,1])(x) == x**2*exp(-x/4)/128


def test_GeneralizedMultivariateLogGammaDistribution():
    h = S.Half
    omega = Matrix([[1, h, h, h],
                     [h, 1, h, h],
                     [h, h, 1, h],
                     [h, h, h, 1]])
    v, l, mu = (4, [1, 2, 3, 4], [1, 2, 3, 4])
    y_1, y_2, y_3, y_4 = symbols('y_1:5', real=True)
    delta = symbols('d', positive=True)
    G = GMVLGO('G', omega, v, l, mu)
    Gd = GMVLG('Gd', delta, v, l, mu)
    dend = ("d**4*Sum(4*24**(-n - 4)*(1 - d)**n*exp((n + 4)*(y_1 + 2*y_2 + 3*y_3 "
            "+ 4*y_4) - exp(y_1) - exp(2*y_2)/2 - exp(3*y_3)/3 - exp(4*y_4)/4)/"
            "(gamma(n + 1)*gamma(n + 4)**3), (n, 0, oo))")
    assert str(density(Gd)(y_1, y_2, y_3, y_4)) == dend
    den = ("5*2**(2/3)*5**(1/3)*Sum(4*24**(-n - 4)*(-2**(2/3)*5**(1/3)/4 + 1)**n*"
          "exp((n + 4)*(y_1 + 2*y_2 + 3*y_3 + 4*y_4) - exp(y_1) - exp(2*y_2)/2 - "
          "exp(3*y_3)/3 - exp(4*y_4)/4)/(gamma(n + 1)*gamma(n + 4)**3), (n, 0, oo))/64")
    assert str(density(G)(y_1, y_2, y_3, y_4)) == den
    marg = ("5*2**(2/3)*5**(1/3)*exp(4*y_1)*exp(-exp(y_1))*Integral(exp(-exp(4*G[3])"
            "/4)*exp(16*G[3])*Integral(exp(-exp(3*G[2])/3)*exp(12*G[2])*Integral(exp("
            "-exp(2*G[1])/2)*exp(8*G[1])*Sum((-1/4)**n*(-4 + 2**(2/3)*5**(1/3"
            "))**n*exp(n*y_1)*exp(2*n*G[1])*exp(3*n*G[2])*exp(4*n*G[3])/(24**n*gamma(n + 1)"
            "*gamma(n + 4)**3), (n, 0, oo)), (G[1], -oo, oo)), (G[2], -oo, oo)), (G[3]"
            ", -oo, oo))/5308416")
    assert str(marginal_distribution(G, G[0])(y_1)) == marg
    omega_f1 = Matrix([[1, h, h]])
    omega_f2 = Matrix([[1, h, h, h],
                     [h, 1, 2, h],
                     [h, h, 1, h],
                     [h, h, h, 1]])
    omega_f3 = Matrix([[6, h, h, h],
                     [h, 1, 2, h],
                     [h, h, 1, h],
                     [h, h, h, 1]])
    v_f = symbols("v_f", positive=False, real=True)
    l_f = [1, 2, v_f, 4]
    m_f = [v_f, 2, 3, 4]
    omega_f4 = Matrix([[1, h, h, h, h],
                     [h, 1, h, h, h],
                     [h, h, 1, h, h],
                     [h, h, h, 1, h],
                     [h, h, h, h, 1]])
    l_f1 = [1, 2, 3, 4, 5]
    omega_f5 = Matrix([[1]])
    mu_f5 = l_f5 = [1]

    raises(ValueError, lambda: GMVLGO('G', omega_f1, v, l, mu))
    raises(ValueError, lambda: GMVLGO('G', omega_f2, v, l, mu))
    raises(ValueError, lambda: GMVLGO('G', omega_f3, v, l, mu))
    raises(ValueError, lambda: GMVLGO('G', omega, v_f, l, mu))
    raises(ValueError, lambda: GMVLGO('G', omega, v, l_f, mu))
    raises(ValueError, lambda: GMVLGO('G', omega, v, l, m_f))
    raises(ValueError, lambda: GMVLGO('G', omega_f4, v, l, mu))
    raises(ValueError, lambda: GMVLGO('G', omega, v, l_f1, mu))
    raises(ValueError, lambda: GMVLGO('G', omega_f5, v, l_f5, mu_f5))
    raises(ValueError, lambda: GMVLG('G', Rational(3, 2), v, l, mu))


def test_MultivariateBeta():
    a1, a2 = symbols('a1, a2', positive=True)
    a1_f, a2_f = symbols('a1, a2', positive=False, real=True)
    mb = MultivariateBeta('B', [a1, a2])
    mb_c = MultivariateBeta('C', a1, a2)
    assert density(mb)(1, 2) == S(2)**(a2 - 1)*gamma(a1 + a2)/\
                                (gamma(a1)*gamma(a2))
    assert marginal_distribution(mb_c, 0)(3) == S(3)**(a1 - 1)*gamma(a1 + a2)/\
                                                (a2*gamma(a1)*gamma(a2))
    raises(ValueError, lambda: MultivariateBeta('b1', [a1_f, a2]))
    raises(ValueError, lambda: MultivariateBeta('b2', [a1, a2_f]))
    raises(ValueError, lambda: MultivariateBeta('b3', [0, 0]))
    raises(ValueError, lambda: MultivariateBeta('b4', [a1_f, a2_f]))
    assert mb.pspace.distribution.set == ProductSet(Interval(0, 1), Interval(0, 1))


def test_MultivariateEwens():
    n, theta, i = symbols('n theta i', positive=True)

    # tests for integer dimensions
    theta_f = symbols('t_f', negative=True)
    a = symbols('a_1:4', positive = True, integer = True)
    ed = MultivariateEwens('E', 3, theta)
    assert density(ed)(a[0], a[1], a[2]) == Piecewise((6*2**(-a[1])*3**(-a[2])*
                                            theta**a[0]*theta**a[1]*theta**a[2]/
                                            (theta*(theta + 1)*(theta + 2)*
                                            factorial(a[0])*factorial(a[1])*
                                            factorial(a[2])), Eq(a[0] + 2*a[1] +
                                            3*a[2], 3)), (0, True))
    assert marginal_distribution(ed, ed[1])(a[1]) == Piecewise((6*2**(-a[1])*
                                                    theta**a[1]/((theta + 1)*
                                                    (theta + 2)*factorial(a[1])),
                                                    Eq(2*a[1] + 1, 3)), (0, True))
    raises(ValueError, lambda: MultivariateEwens('e1', 5, theta_f))
    assert ed.pspace.distribution.set == ProductSet(Range(0, 4, 1),
                                            Range(0, 2, 1), Range(0, 2, 1))

    # tests for symbolic dimensions
    eds = MultivariateEwens('E', n, theta)
    a = IndexedBase('a')
    j, k = symbols('j, k')
    den = Piecewise((factorial(n)*Product(theta**a[j]*(j + 1)**(-a[j])/
           factorial(a[j]), (j, 0, n - 1))/RisingFactorial(theta, n),
            Eq(n, Sum((k + 1)*a[k], (k, 0, n - 1)))), (0, True))
    assert density(eds)(a).dummy_eq(den)


def test_Multinomial():
    n, x1, x2, x3, x4 = symbols('n, x1, x2, x3, x4', nonnegative=True, integer=True)
    p1, p2, p3, p4 = symbols('p1, p2, p3, p4', positive=True)
    p1_f, n_f = symbols('p1_f, n_f', negative=True)
    M = Multinomial('M', n, [p1, p2, p3, p4])
    C = Multinomial('C', 3, p1, p2, p3)
    f = factorial
    assert density(M)(x1, x2, x3, x4) == Piecewise((p1**x1*p2**x2*p3**x3*p4**x4*
                                            f(n)/(f(x1)*f(x2)*f(x3)*f(x4)),
                                            Eq(n, x1 + x2 + x3 + x4)), (0, True))
    assert marginal_distribution(C, C[0])(x1).subs(x1, 1) ==\
                                                            3*p1*p2**2 +\
                                                            6*p1*p2*p3 +\
                                                            3*p1*p3**2
    raises(ValueError, lambda: Multinomial('b1', 5, [p1, p2, p3, p1_f]))
    raises(ValueError, lambda: Multinomial('b2', n_f, [p1, p2, p3, p4]))
    raises(ValueError, lambda: Multinomial('b3', n, 0.5, 0.4, 0.3, 0.1))


def test_NegativeMultinomial():
    k0, x1, x2, x3, x4 = symbols('k0, x1, x2, x3, x4', nonnegative=True, integer=True)
    p1, p2, p3, p4 = symbols('p1, p2, p3, p4', positive=True)
    p1_f = symbols('p1_f', negative=True)
    N = NegativeMultinomial('N', 4, [p1, p2, p3, p4])
    C = NegativeMultinomial('C', 4, 0.1, 0.2, 0.3)
    g = gamma
    f = factorial
    assert simplify(density(N)(x1, x2, x3, x4) -
            p1**x1*p2**x2*p3**x3*p4**x4*(-p1 - p2 - p3 - p4 + 1)**4*g(x1 + x2 +
            x3 + x4 + 4)/(6*f(x1)*f(x2)*f(x3)*f(x4))) is S.Zero
    assert comp(marginal_distribution(C, C[0])(1).evalf(), 0.33, .01)
    raises(ValueError, lambda: NegativeMultinomial('b1', 5, [p1, p2, p3, p1_f]))
    raises(ValueError, lambda: NegativeMultinomial('b2', k0, 0.5, 0.4, 0.3, 0.4))
    assert N.pspace.distribution.set == ProductSet(Range(0, oo, 1),
                    Range(0, oo, 1), Range(0, oo, 1), Range(0, oo, 1))


@slow
def test_JointPSpace_marginal_distribution():
    T = MultivariateT('T', [0, 0], [[1, 0], [0, 1]], 2)
    got = marginal_distribution(T, T[1])(x)
    ans = sqrt(2)*(x**2/2 + 1)/(4*polar_lift(x**2/2 + 1)**(S(5)/2))
    assert got == ans, got
    assert integrate(marginal_distribution(T, 1)(x), (x, -oo, oo)) == 1

    t = MultivariateT('T', [0, 0, 0], [[1, 0, 0], [0, 1, 0], [0, 0, 1]], 3)
    assert comp(marginal_distribution(t, 0)(1).evalf(), 0.2, .01)


def test_JointRV():
    x1, x2 = (Indexed('x', i) for i in (1, 2))
    pdf = exp(-x1**2/2 + x1 - x2**2/2 - S.Half)/(2*pi)
    X = JointRV('x', pdf)
    assert density(X)(1, 2) == exp(-2)/(2*pi)
    assert isinstance(X.pspace.distribution, JointDistributionHandmade)
    assert marginal_distribution(X, 0)(2) == sqrt(2)*exp(Rational(-1, 2))/(2*sqrt(pi))


def test_expectation():
    m = Normal('A', [x, y], [[1, 0], [0, 1]])
    assert simplify(E(m[1])) == y


@XFAIL
def test_joint_vector_expectation():
    m = Normal('A', [x, y], [[1, 0], [0, 1]])
    assert E(m) == (x, y)


def test_sample_numpy():
    distribs_numpy = [
        MultivariateNormal("M", [3, 4], [[2, 1], [1, 2]]),
        MultivariateBeta("B", [0.4, 5, 15, 50, 203]),
        Multinomial("N", 50, [0.3, 0.2, 0.1, 0.25, 0.15])
    ]
    size = 3
    numpy = import_module('numpy')
    if not numpy:
        skip('Numpy is not installed. Abort tests for _sample_numpy.')
    else:
        for X in distribs_numpy:
            samps = sample(X, size=size, library='numpy')
            for sam in samps:
                assert tuple(sam) in X.pspace.distribution.set
        N_c = NegativeMultinomial('N', 3, 0.1, 0.1, 0.1)
        raises(NotImplementedError, lambda: sample(N_c, library='numpy'))


def test_sample_scipy():
    distribs_scipy = [
        MultivariateNormal("M", [0, 0], [[0.1, 0.025], [0.025, 0.1]]),
        MultivariateBeta("B", [0.4, 5, 15]),
        Multinomial("N", 8, [0.3, 0.2, 0.1, 0.4])
    ]

    size = 3
    scipy = import_module('scipy')
    if not scipy:
        skip('Scipy not installed. Abort tests for _sample_scipy.')
    else:
        for X in distribs_scipy:
            samps = sample(X, size=size)
            samps2 = sample(X, size=(2, 2))
            for sam in samps:
                assert tuple(sam) in X.pspace.distribution.set
            for i in range(2):
                for j in range(2):
                    assert tuple(samps2[i][j]) in X.pspace.distribution.set
        N_c = NegativeMultinomial('N', 3, 0.1, 0.1, 0.1)
        raises(NotImplementedError, lambda: sample(N_c))


def test_sample_pymc():
    distribs_pymc = [
        MultivariateNormal("M", [5, 2], [[1, 0], [0, 1]]),
        MultivariateBeta("B", [0.4, 5, 15]),
        Multinomial("N", 4, [0.3, 0.2, 0.1, 0.4])
    ]
    size = 3
    pymc = import_module('pymc')
    if not pymc:
        skip('PyMC is not installed. Abort tests for _sample_pymc.')
    else:
        for X in distribs_pymc:
            samps = sample(X, size=size, library='pymc')
            for sam in samps:
                assert tuple(sam.flatten()) in X.pspace.distribution.set
        N_c = NegativeMultinomial('N', 3, 0.1, 0.1, 0.1)
        raises(NotImplementedError, lambda: sample(N_c, library='pymc'))


def test_sample_seed():
    x1, x2 = (Indexed('x', i) for i in (1, 2))
    pdf = exp(-x1**2/2 + x1 - x2**2/2 - S.Half)/(2*pi)
    X = JointRV('x', pdf)

    libraries = ['scipy', 'numpy', 'pymc']
    for lib in libraries:
        try:
            imported_lib = import_module(lib)
            if imported_lib:
                s0, s1, s2 = [], [], []
                s0 = sample(X, size=10, library=lib, seed=0)
                s1 = sample(X, size=10, library=lib, seed=0)
                s2 = sample(X, size=10, library=lib, seed=1)
                assert all(s0 == s1)
                assert all(s1 != s2)
        except NotImplementedError:
            continue

#
# XXX: This fails for pymc. Previously the test appeared to pass but that is
# just because the library argument was not passed so the test always used
# scipy.
#
def test_issue_21057():
    m = Normal("x", [0, 0], [[0, 0], [0, 0]])
    n = MultivariateNormal("x", [0, 0], [[0, 0], [0, 0]])
    p = Normal("x", [0, 0], [[0, 0], [0, 1]])
    assert m == n
    libraries = ('scipy', 'numpy')  # , 'pymc')  # <-- pymc fails
    for library in libraries:
        try:
            imported_lib = import_module(library)
            if imported_lib:
                s1 = sample(m, size=8, library=library)
                s2 = sample(n, size=8, library=library)
                s3 = sample(p, size=8, library=library)
                assert tuple(s1.flatten()) == tuple(s2.flatten())
                for s in s3:
                    assert tuple(s.flatten()) in p.pspace.distribution.set
        except NotImplementedError:
            continue


#
# When this passes the pymc part can be uncommented in test_issue_21057 above
# and this can be deleted.
#
@XFAIL
def test_issue_21057_pymc():
    m = Normal("x", [0, 0], [[0, 0], [0, 0]])
    n = MultivariateNormal("x", [0, 0], [[0, 0], [0, 0]])
    p = Normal("x", [0, 0], [[0, 0], [0, 1]])
    assert m == n
    libraries = ('pymc',)
    for library in libraries:
        try:
            imported_lib = import_module(library)
            if imported_lib:
                s1 = sample(m, size=8, library=library)
                s2 = sample(n, size=8, library=library)
                s3 = sample(p, size=8, library=library)
                assert tuple(s1.flatten()) == tuple(s2.flatten())
                for s in s3:
                    assert tuple(s.flatten()) in p.pspace.distribution.set
        except NotImplementedError:
            continue

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\bs4\tests\test_dammit.py ===
# encoding: utf-8
import pytest
import logging
import warnings
import bs4
from bs4 import BeautifulSoup
from bs4.dammit import (
    EntitySubstitution,
    EncodingDetector,
    UnicodeDammit,
)


class TestUnicodeDammit(object):
    """Standalone tests of UnicodeDammit."""

    def test_unicode_input(self):
        markup = "I'm already Unicode! \N{SNOWMAN}"
        dammit = UnicodeDammit(markup)
        assert dammit.unicode_markup == markup

    @pytest.mark.parametrize(
        "smart_quotes_to,expect_converted",
        [
            (None, "\u2018\u2019\u201c\u201d"),
            ("xml", "&#x2018;&#x2019;&#x201C;&#x201D;"),
            ("html", "&lsquo;&rsquo;&ldquo;&rdquo;"),
            ("ascii", "''" + '""'),
        ],
    )
    def test_smart_quotes_to(self, smart_quotes_to, expect_converted):
        """Verify the functionality of the smart_quotes_to argument
        to the UnicodeDammit constructor."""
        markup = b"<foo>\x91\x92\x93\x94</foo>"
        converted = UnicodeDammit(
            markup,
            known_definite_encodings=["windows-1252"],
            smart_quotes_to=smart_quotes_to,
        ).unicode_markup
        assert converted == "<foo>{}</foo>".format(expect_converted)

    def test_detect_utf8(self):
        utf8 = b"Sacr\xc3\xa9 bleu! \xe2\x98\x83"
        dammit = UnicodeDammit(utf8)
        assert dammit.original_encoding.lower() == "utf-8"
        assert dammit.unicode_markup == "Sacr\xe9 bleu! \N{SNOWMAN}"

    def test_convert_hebrew(self):
        hebrew = b"\xed\xe5\xec\xf9"
        dammit = UnicodeDammit(hebrew, ["iso-8859-8"])
        assert dammit.original_encoding.lower() == "iso-8859-8"
        assert dammit.unicode_markup == "\u05dd\u05d5\u05dc\u05e9"

    def test_dont_see_smart_quotes_where_there_are_none(self):
        utf_8 = b"\343\202\261\343\203\274\343\202\277\343\202\244 Watch"
        dammit = UnicodeDammit(utf_8)
        assert dammit.original_encoding.lower() == "utf-8"
        assert dammit.unicode_markup.encode("utf-8") == utf_8

    def test_ignore_inappropriate_codecs(self):
        utf8_data = "Räksmörgås".encode("utf-8")
        dammit = UnicodeDammit(utf8_data, ["iso-8859-8"])
        assert dammit.original_encoding.lower() == "utf-8"

    def test_ignore_invalid_codecs(self):
        utf8_data = "Räksmörgås".encode("utf-8")
        for bad_encoding in [".utf8", "...", "utF---16.!"]:
            dammit = UnicodeDammit(utf8_data, [bad_encoding])
            assert dammit.original_encoding.lower() == "utf-8"

    def test_exclude_encodings(self):
        # This is UTF-8.
        utf8_data = "Räksmörgås".encode("utf-8")

        # But if we exclude UTF-8 from consideration, the guess is
        # Windows-1252.
        dammit = UnicodeDammit(utf8_data, exclude_encodings=["utf-8"])
        assert dammit.original_encoding.lower() == "windows-1252"

        # And if we exclude that, there is no valid guess at all.
        dammit = UnicodeDammit(utf8_data, exclude_encodings=["utf-8", "windows-1252"])
        assert dammit.original_encoding is None


class TestEncodingDetector(object):
    def test_encoding_detector_replaces_junk_in_encoding_name_with_replacement_character(
        self,
    ):
        detected = EncodingDetector(b'<?xml version="1.0" encoding="UTF-\xdb" ?>')
        encodings = list(detected.encodings)
        assert "utf-\N{REPLACEMENT CHARACTER}" in encodings

    def test_detect_html5_style_meta_tag(self):
        for data in (
            b'<html><meta charset="euc-jp" /></html>',
            b"<html><meta charset='euc-jp' /></html>",
            b"<html><meta charset=euc-jp /></html>",
            b"<html><meta charset=euc-jp/></html>",
        ):
            dammit = UnicodeDammit(data, is_html=True)
            assert "euc-jp" == dammit.original_encoding

    def test_last_ditch_entity_replacement(self):
        # This is a UTF-8 document that contains bytestrings
        # completely incompatible with UTF-8 (ie. encoded with some other
        # encoding).
        #
        # Since there is no consistent encoding for the document,
        # Unicode, Dammit will eventually encode the document as UTF-8
        # and encode the incompatible characters as REPLACEMENT
        # CHARACTER.
        #
        # If chardet is installed, it will detect that the document
        # can be converted into ISO-8859-1 without errors. This happens
        # to be the wrong encoding, but it is a consistent encoding, so the
        # code we're testing here won't run.
        #
        # So we temporarily disable chardet if it's present.
        doc = b"""\357\273\277<?xml version="1.0" encoding="UTF-8"?>
<html><b>\330\250\330\252\330\261</b>
<i>\310\322\321\220\312\321\355\344</i></html>"""
        chardet = bs4.dammit._chardet_dammit
        logging.disable(logging.WARNING)
        try:

            def noop(str):
                return None

            bs4.dammit._chardet_dammit = noop
            dammit = UnicodeDammit(doc)
            assert True is dammit.contains_replacement_characters
            assert "\ufffd" in dammit.unicode_markup

            soup = BeautifulSoup(doc, "html.parser")
            assert soup.contains_replacement_characters
        finally:
            logging.disable(logging.NOTSET)
            bs4.dammit._chardet_dammit = chardet

    def test_byte_order_mark_removed(self):
        # A document written in UTF-16LE will have its byte order marker stripped.
        data = b"\xff\xfe<\x00a\x00>\x00\xe1\x00\xe9\x00<\x00/\x00a\x00>\x00"
        dammit = UnicodeDammit(data)
        assert "<a>áé</a>" == dammit.unicode_markup
        assert "utf-16le" == dammit.original_encoding

    def test_known_definite_versus_user_encodings(self):
        # The known_definite_encodings are used before sniffing the
        # byte-order mark; the user_encodings are used afterwards.

        # Here's a document in UTF-16LE.
        data = b"\xff\xfe<\x00a\x00>\x00\xe1\x00\xe9\x00<\x00/\x00a\x00>\x00"
        dammit = UnicodeDammit(data)

        # We can process it as UTF-16 by passing it in as a known
        # definite encoding.
        before = UnicodeDammit(data, known_definite_encodings=["utf-16"])
        assert "utf-16" == before.original_encoding

        # If we pass UTF-18 as a user encoding, it's not even
        # tried--the encoding sniffed from the byte-order mark takes
        # precedence.
        after = UnicodeDammit(data, user_encodings=["utf-8"])
        assert "utf-16le" == after.original_encoding
        assert ["utf-16le"] == [x[0] for x in dammit.tried_encodings]

        # Here's a document in ISO-8859-8.
        hebrew = b"\xed\xe5\xec\xf9"
        dammit = UnicodeDammit(
            hebrew, known_definite_encodings=["utf-8"], user_encodings=["iso-8859-8"]
        )

        # The known_definite_encodings don't work, BOM sniffing does
        # nothing (it only works for a few UTF encodings), but one of
        # the user_encodings does work.
        assert "iso-8859-8" == dammit.original_encoding
        assert ["utf-8", "iso-8859-8"] == [x[0] for x in dammit.tried_encodings]

    def test_deprecated_override_encodings(self):
        # override_encodings is a deprecated alias for
        # known_definite_encodings.
        hebrew = b"\xed\xe5\xec\xf9"
        with warnings.catch_warnings(record=True) as w:
            dammit = UnicodeDammit(
                hebrew,
                known_definite_encodings=["shift-jis"],
                override_encodings=["utf-8"],
                user_encodings=["iso-8859-8"],
            )
        [warning] = w
        message = warning.message
        assert isinstance(message, DeprecationWarning)
        assert warning.filename == __file__
        assert "iso-8859-8" == dammit.original_encoding

        # known_definite_encodings and override_encodings were tried
        # before user_encodings.
        assert ["shift-jis", "utf-8", "iso-8859-8"] == (
            [x[0] for x in dammit.tried_encodings]
        )

    def test_detwingle(self):
        # Here's a UTF8 document.
        utf8 = ("\N{SNOWMAN}" * 3).encode("utf8")

        # Here's a Windows-1252 document.
        windows_1252 = (
            "\N{LEFT DOUBLE QUOTATION MARK}Hi, I like Windows!"
            "\N{RIGHT DOUBLE QUOTATION MARK}"
        ).encode("windows_1252")

        # Through some unholy alchemy, they've been stuck together.
        doc = utf8 + windows_1252 + utf8

        # The document can't be turned into UTF-8:
        with pytest.raises(UnicodeDecodeError):
            doc.decode("utf8")

        # Unicode, Dammit thinks the whole document is Windows-1252,
        # and decodes it into "â˜ƒâ˜ƒâ˜ƒ“Hi, I like Windows!”â˜ƒâ˜ƒâ˜ƒ"

        # But if we run it through fix_embedded_windows_1252, it's fixed:
        fixed = UnicodeDammit.detwingle(doc)
        assert "☃☃☃“Hi, I like Windows!”☃☃☃" == fixed.decode("utf8")

    def test_detwingle_ignores_multibyte_characters(self):
        # Each of these characters has a UTF-8 representation ending
        # in \x93. \x93 is a smart quote if interpreted as
        # Windows-1252. But our code knows to skip over multibyte
        # UTF-8 characters, so they'll survive the process unscathed.
        for tricky_unicode_char in (
            "\N{LATIN SMALL LIGATURE OE}",  # 2-byte char '\xc5\x93'
            "\N{LATIN SUBSCRIPT SMALL LETTER X}",  # 3-byte char '\xe2\x82\x93'
            "\xf0\x90\x90\x93",  # This is a CJK character, not sure which one.
        ):
            input = tricky_unicode_char.encode("utf8")
            assert input.endswith(b"\x93")
            output = UnicodeDammit.detwingle(input)
            assert output == input

    def test_find_declared_encoding(self):
        # Test our ability to find a declared encoding inside an
        # XML or HTML document.
        #
        # Even if the document comes in as Unicode, it may be
        # interesting to know what encoding was claimed
        # originally.

        html_unicode = '<html><head><meta charset="utf-8"></head></html>'
        html_bytes = html_unicode.encode("ascii")

        xml_unicode = '<?xml version="1.0" encoding="ISO-8859-1" ?>'
        xml_bytes = xml_unicode.encode("ascii")

        m = EncodingDetector.find_declared_encoding
        assert m(html_unicode, is_html=False) is None
        assert "utf-8" == m(html_unicode, is_html=True)
        assert "utf-8" == m(html_bytes, is_html=True)

        assert "iso-8859-1" == m(xml_unicode)
        assert "iso-8859-1" == m(xml_bytes)

        # Normally, only the first few kilobytes of a document are checked for
        # an encoding.
        spacer = b" " * 5000
        assert m(spacer + html_bytes) is None
        assert m(spacer + xml_bytes) is None

        # But you can tell find_declared_encoding to search an entire
        # HTML document.
        assert (
            m(spacer + html_bytes, is_html=True, search_entire_document=True) == "utf-8"
        )

        # The XML encoding declaration has to be the very first thing
        # in the document. We'll allow whitespace before the document
        # starts, but nothing else.
        assert m(xml_bytes, search_entire_document=True) == "iso-8859-1"
        assert m(b" " + xml_bytes, search_entire_document=True) == "iso-8859-1"
        assert m(b"a" + xml_bytes, search_entire_document=True) is None


class TestEntitySubstitution(object):
    """Standalone tests of the EntitySubstitution class."""

    def setup_method(self):
        self.sub = EntitySubstitution

    @pytest.mark.parametrize(
        "original,substituted",
        [
            # Basic case. Unicode characters corresponding to named
            # HTML entites are substituted; others are not.
            ("foo\u2200\N{SNOWMAN}\u00f5bar", "foo&forall;\N{SNOWMAN}&otilde;bar"),
            # MS smart quotes are a common source of frustration, so we
            # give them a special test.
            ("‘’foo“”", "&lsquo;&rsquo;foo&ldquo;&rdquo;"),
        ],
    )
    def test_substitute_html(self, original, substituted):
        assert self.sub.substitute_html(original) == substituted

    def test_html5_entity(self):
        for entity, u in (
            # A few spot checks of our ability to recognize
            # special character sequences and convert them
            # to named entities.
            ("&models;", "\u22a7"),
            ("&Nfr;", "\U0001d511"),
            ("&ngeqq;", "\u2267\u0338"),
            ("&not;", "\xac"),
            ("&Not;", "\u2aec"),
            # We _could_ convert | to &verbarr;, but we don't, because
            # | is an ASCII character.
            ("|" "|"),
            # Similarly for the fj ligature, which we could convert to
            # &fjlig;, but we don't.
            ("fj", "fj"),
            # We do convert _these_ ASCII characters to HTML entities,
            # because that's required to generate valid HTML.
            ("&gt;", ">"),
            ("&lt;", "<"),
        ):
            template = "3 %s 4"
            raw = template % u
            with_entities = template % entity
            assert self.sub.substitute_html(raw) == with_entities

    def test_html5_entity_with_variation_selector(self):
        # Some HTML5 entities correspond either to a single-character
        # Unicode sequence _or_ to the same character plus U+FE00,
        # VARIATION SELECTOR 1. We can handle this.
        data = "fjords \u2294 penguins"
        markup = "fjords &sqcup; penguins"
        assert self.sub.substitute_html(data) == markup

        data = "fjords \u2294\ufe00 penguins"
        markup = "fjords &sqcups; penguins"
        assert self.sub.substitute_html(data) == markup

    def test_xml_converstion_includes_no_quotes_if_make_quoted_attribute_is_false(self):
        s = 'Welcome to "my bar"'
        assert self.sub.substitute_xml(s, False) == s

    def test_xml_attribute_quoting_normally_uses_double_quotes(self):
        assert self.sub.substitute_xml("Welcome", True) == '"Welcome"'
        assert self.sub.substitute_xml("Bob's Bar", True) == '"Bob\'s Bar"'

    def test_xml_attribute_quoting_uses_single_quotes_when_value_contains_double_quotes(
        self,
    ):
        s = 'Welcome to "my bar"'
        assert self.sub.substitute_xml(s, True) == "'Welcome to \"my bar\"'"

    def test_xml_attribute_quoting_escapes_single_quotes_when_value_contains_both_single_and_double_quotes(
        self,
    ):
        s = 'Welcome to "Bob\'s Bar"'
        assert self.sub.substitute_xml(s, True) == '"Welcome to &quot;Bob\'s Bar&quot;"'

    def test_xml_quotes_arent_escaped_when_value_is_not_being_quoted(self):
        quoted = 'Welcome to "Bob\'s Bar"'
        assert self.sub.substitute_xml(quoted) == quoted

    def test_xml_quoting_handles_angle_brackets(self):
        assert self.sub.substitute_xml("foo<bar>") == "foo&lt;bar&gt;"

    def test_xml_quoting_handles_ampersands(self):
        assert self.sub.substitute_xml("AT&T") == "AT&amp;T"

    def test_xml_quoting_including_ampersands_when_they_are_part_of_an_entity(self):
        assert self.sub.substitute_xml("&Aacute;T&T") == "&amp;Aacute;T&amp;T"

    def test_xml_quoting_ignoring_ampersands_when_they_are_part_of_an_entity(self):
        assert (
            self.sub.substitute_xml_containing_entities("&Aacute;T&T")
            == "&Aacute;T&amp;T"
        )

    def test_quotes_not_html_substituted(self):
        """There's no need to do this except inside attribute values."""
        text = 'Bob\'s "bar"'
        assert self.sub.substitute_html(text) == text

    @pytest.mark.parametrize(
        "markup, old",
        [
            ("foo & bar", "foo &amp; bar"),
            ("foo&", "foo&amp;"),
            ("foo&&& bar", "foo&amp;&amp;&amp; bar"),
            ("x=1&y=2", "x=1&amp;y=2"),
            ("&123", "&amp;123"),
            ("&abc", "&amp;abc"),
            ("foo &0 bar", "foo &amp;0 bar"),
            ("foo &lolwat bar", "foo &amp;lolwat bar"),
        ],
    )
    def test_unambiguous_ampersands_not_escaped(self, markup, old):
        assert self.sub.substitute_html(markup) == old
        assert self.sub.substitute_html5_raw(markup) == markup

    @pytest.mark.parametrize(
        "markup,html,html5,html5raw",
        [
            ("&divide;", "&amp;divide;", "&amp;divide;", "&divide;"),
            ("&nonesuch;", "&amp;nonesuch;", "&amp;nonesuch;", "&amp;nonesuch;"),
            ("&#247;", "&amp;#247;", "&amp;#247;", "&amp;#247;"),
            ("&#xa1;", "&amp;#xa1;", "&amp;#xa1;", "&amp;#xa1;"),
        ],
    )
    def test_when_entity_ampersands_are_escaped(self, markup, html, html5, html5raw):
        # The html and html5 formatters always escape the ampersand
        # that begins an entity reference, because they assume
        # Beautiful Soup has already converted any unescaped entity references
        # to Unicode characters.
        #
        # The html5_raw formatter does not escape the ampersand that
        # begins a recognized HTML entity, because it does not
        # fit the HTML5 definition of an ambiguous ampersand.
        #
        # The html5_raw formatter does escape the ampersands in front
        # of unrecognized named entities, as well as numeric and
        # hexadecimal entities, because they do fit the definition.
        assert self.sub.substitute_html(markup) == html
        assert self.sub.substitute_html5(markup) == html5
        assert self.sub.substitute_html5_raw(markup) == html5raw

    @pytest.mark.parametrize(
        "markup,expect", [("&nosuchentity;", "&amp;nosuchentity;")]
    )
    def test_ambiguous_ampersands_escaped(self, markup, expect):
        assert self.sub.substitute_html(markup) == expect
        assert self.sub.substitute_html5_raw(markup) == expect

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\copy_view\test_interp_fillna.py ===
import numpy as np
import pytest

from pandas import (
    NA,
    ArrowDtype,
    DataFrame,
    Interval,
    NaT,
    Series,
    Timestamp,
    interval_range,
    option_context,
)
import pandas._testing as tm
from pandas.tests.copy_view.util import get_array


@pytest.mark.parametrize("method", ["pad", "nearest", "linear"])
def test_interpolate_no_op(using_copy_on_write, method):
    df = DataFrame({"a": [1, 2]})
    df_orig = df.copy()

    warn = None
    if method == "pad":
        warn = FutureWarning
    msg = "DataFrame.interpolate with method=pad is deprecated"
    with tm.assert_produces_warning(warn, match=msg):
        result = df.interpolate(method=method)

    if using_copy_on_write:
        assert np.shares_memory(get_array(result, "a"), get_array(df, "a"))
    else:
        assert not np.shares_memory(get_array(result, "a"), get_array(df, "a"))

    result.iloc[0, 0] = 100

    if using_copy_on_write:
        assert not np.shares_memory(get_array(result, "a"), get_array(df, "a"))
    tm.assert_frame_equal(df, df_orig)


@pytest.mark.parametrize("func", ["ffill", "bfill"])
def test_interp_fill_functions(using_copy_on_write, func):
    # Check that these takes the same code paths as interpolate
    df = DataFrame({"a": [1, 2]})
    df_orig = df.copy()

    result = getattr(df, func)()

    if using_copy_on_write:
        assert np.shares_memory(get_array(result, "a"), get_array(df, "a"))
    else:
        assert not np.shares_memory(get_array(result, "a"), get_array(df, "a"))

    result.iloc[0, 0] = 100

    if using_copy_on_write:
        assert not np.shares_memory(get_array(result, "a"), get_array(df, "a"))
    tm.assert_frame_equal(df, df_orig)


@pytest.mark.parametrize("func", ["ffill", "bfill"])
@pytest.mark.parametrize(
    "vals", [[1, np.nan, 2], [Timestamp("2019-12-31"), NaT, Timestamp("2020-12-31")]]
)
def test_interpolate_triggers_copy(using_copy_on_write, vals, func):
    df = DataFrame({"a": vals})
    result = getattr(df, func)()

    assert not np.shares_memory(get_array(result, "a"), get_array(df, "a"))
    if using_copy_on_write:
        # Check that we don't have references when triggering a copy
        assert result._mgr._has_no_reference(0)


@pytest.mark.parametrize(
    "vals", [[1, np.nan, 2], [Timestamp("2019-12-31"), NaT, Timestamp("2020-12-31")]]
)
def test_interpolate_inplace_no_reference_no_copy(using_copy_on_write, vals):
    df = DataFrame({"a": vals})
    arr = get_array(df, "a")
    df.interpolate(method="linear", inplace=True)

    assert np.shares_memory(arr, get_array(df, "a"))
    if using_copy_on_write:
        # Check that we don't have references when triggering a copy
        assert df._mgr._has_no_reference(0)


@pytest.mark.parametrize(
    "vals", [[1, np.nan, 2], [Timestamp("2019-12-31"), NaT, Timestamp("2020-12-31")]]
)
def test_interpolate_inplace_with_refs(using_copy_on_write, vals, warn_copy_on_write):
    df = DataFrame({"a": [1, np.nan, 2]})
    df_orig = df.copy()
    arr = get_array(df, "a")
    view = df[:]
    with tm.assert_cow_warning(warn_copy_on_write):
        df.interpolate(method="linear", inplace=True)

    if using_copy_on_write:
        # Check that copy was triggered in interpolate and that we don't
        # have any references left
        assert not np.shares_memory(arr, get_array(df, "a"))
        tm.assert_frame_equal(df_orig, view)
        assert df._mgr._has_no_reference(0)
        assert view._mgr._has_no_reference(0)
    else:
        assert np.shares_memory(arr, get_array(df, "a"))


@pytest.mark.parametrize("func", ["ffill", "bfill"])
@pytest.mark.parametrize("dtype", ["float64", "Float64"])
def test_interp_fill_functions_inplace(
    using_copy_on_write, func, warn_copy_on_write, dtype
):
    # Check that these takes the same code paths as interpolate
    df = DataFrame({"a": [1, np.nan, 2]}, dtype=dtype)
    df_orig = df.copy()
    arr = get_array(df, "a")
    view = df[:]

    with tm.assert_cow_warning(warn_copy_on_write and dtype == "float64"):
        getattr(df, func)(inplace=True)

    if using_copy_on_write:
        # Check that copy was triggered in interpolate and that we don't
        # have any references left
        assert not np.shares_memory(arr, get_array(df, "a"))
        tm.assert_frame_equal(df_orig, view)
        assert df._mgr._has_no_reference(0)
        assert view._mgr._has_no_reference(0)
    else:
        assert np.shares_memory(arr, get_array(df, "a")) is (dtype == "float64")


def test_interpolate_cannot_with_object_dtype(using_copy_on_write):
    df = DataFrame({"a": ["a", np.nan, "c"], "b": 1})
    df["a"] = df["a"].astype(object)
    df_orig = df.copy()

    msg = "DataFrame.interpolate with object dtype"
    with tm.assert_produces_warning(FutureWarning, match=msg):
        result = df.interpolate(method="linear")

    if using_copy_on_write:
        assert np.shares_memory(get_array(result, "a"), get_array(df, "a"))
    else:
        assert not np.shares_memory(get_array(result, "a"), get_array(df, "a"))

    result.iloc[0, 0] = Timestamp("2021-12-31")

    if using_copy_on_write:
        assert not np.shares_memory(get_array(result, "a"), get_array(df, "a"))
    tm.assert_frame_equal(df, df_orig)


def test_interpolate_object_convert_no_op(using_copy_on_write, using_infer_string):
    df = DataFrame({"a": ["a", "b", "c"], "b": 1})
    df["a"] = df["a"].astype(object)
    arr_a = get_array(df, "a")
    msg = "DataFrame.interpolate with method=pad is deprecated"
    with tm.assert_produces_warning(FutureWarning, match=msg):
        df.interpolate(method="pad", inplace=True)

    # Now CoW makes a copy, it should not!
    if using_copy_on_write and not using_infer_string:
        assert df._mgr._has_no_reference(0)
        assert np.shares_memory(arr_a, get_array(df, "a"))


def test_interpolate_object_convert_copies(using_copy_on_write):
    df = DataFrame({"a": Series([1, 2], dtype=object), "b": 1})
    arr_a = get_array(df, "a")
    msg = "DataFrame.interpolate with method=pad is deprecated"
    with tm.assert_produces_warning(FutureWarning, match=msg):
        df.interpolate(method="pad", inplace=True)

    if using_copy_on_write:
        assert df._mgr._has_no_reference(0)
        assert not np.shares_memory(arr_a, get_array(df, "a"))


def test_interpolate_downcast(using_copy_on_write):
    df = DataFrame({"a": [1, np.nan, 2.5], "b": 1})
    arr_a = get_array(df, "a")
    msg = "DataFrame.interpolate with method=pad is deprecated"
    with tm.assert_produces_warning(FutureWarning, match=msg):
        df.interpolate(method="pad", inplace=True, downcast="infer")

    if using_copy_on_write:
        assert df._mgr._has_no_reference(0)
    assert np.shares_memory(arr_a, get_array(df, "a"))


def test_interpolate_downcast_reference_triggers_copy(using_copy_on_write):
    df = DataFrame({"a": [1, np.nan, 2.5], "b": 1})
    df_orig = df.copy()
    arr_a = get_array(df, "a")
    view = df[:]
    msg = "DataFrame.interpolate with method=pad is deprecated"
    with tm.assert_produces_warning(FutureWarning, match=msg):
        df.interpolate(method="pad", inplace=True, downcast="infer")

    if using_copy_on_write:
        assert df._mgr._has_no_reference(0)
        assert not np.shares_memory(arr_a, get_array(df, "a"))
        tm.assert_frame_equal(df_orig, view)
    else:
        tm.assert_frame_equal(df, view)


def test_fillna(using_copy_on_write):
    df = DataFrame({"a": [1.5, np.nan], "b": 1})
    df_orig = df.copy()

    df2 = df.fillna(5.5)
    if using_copy_on_write:
        assert np.shares_memory(get_array(df, "b"), get_array(df2, "b"))
    else:
        assert not np.shares_memory(get_array(df, "b"), get_array(df2, "b"))

    df2.iloc[0, 1] = 100
    tm.assert_frame_equal(df_orig, df)


def test_fillna_dict(using_copy_on_write):
    df = DataFrame({"a": [1.5, np.nan], "b": 1})
    df_orig = df.copy()

    df2 = df.fillna({"a": 100.5})
    if using_copy_on_write:
        assert np.shares_memory(get_array(df, "b"), get_array(df2, "b"))
        assert not np.shares_memory(get_array(df, "a"), get_array(df2, "a"))
    else:
        assert not np.shares_memory(get_array(df, "b"), get_array(df2, "b"))

    df2.iloc[0, 1] = 100
    tm.assert_frame_equal(df_orig, df)


@pytest.mark.parametrize("downcast", [None, False])
def test_fillna_inplace(using_copy_on_write, downcast):
    df = DataFrame({"a": [1.5, np.nan], "b": 1})
    arr_a = get_array(df, "a")
    arr_b = get_array(df, "b")

    msg = "The 'downcast' keyword in fillna is deprecated"
    with tm.assert_produces_warning(FutureWarning, match=msg):
        df.fillna(5.5, inplace=True, downcast=downcast)
    assert np.shares_memory(get_array(df, "a"), arr_a)
    assert np.shares_memory(get_array(df, "b"), arr_b)
    if using_copy_on_write:
        assert df._mgr._has_no_reference(0)
        assert df._mgr._has_no_reference(1)


def test_fillna_inplace_reference(using_copy_on_write, warn_copy_on_write):
    df = DataFrame({"a": [1.5, np.nan], "b": 1})
    df_orig = df.copy()
    arr_a = get_array(df, "a")
    arr_b = get_array(df, "b")
    view = df[:]

    with tm.assert_cow_warning(warn_copy_on_write):
        df.fillna(5.5, inplace=True)
    if using_copy_on_write:
        assert not np.shares_memory(get_array(df, "a"), arr_a)
        assert np.shares_memory(get_array(df, "b"), arr_b)
        assert view._mgr._has_no_reference(0)
        assert df._mgr._has_no_reference(0)
        tm.assert_frame_equal(view, df_orig)
    else:
        assert np.shares_memory(get_array(df, "a"), arr_a)
        assert np.shares_memory(get_array(df, "b"), arr_b)
    expected = DataFrame({"a": [1.5, 5.5], "b": 1})
    tm.assert_frame_equal(df, expected)


def test_fillna_interval_inplace_reference(using_copy_on_write, warn_copy_on_write):
    # Set dtype explicitly to avoid implicit cast when setting nan
    ser = Series(
        interval_range(start=0, end=5), name="a", dtype="interval[float64, right]"
    )
    ser.iloc[1] = np.nan

    ser_orig = ser.copy()
    view = ser[:]
    with tm.assert_cow_warning(warn_copy_on_write):
        ser.fillna(value=Interval(left=0, right=5), inplace=True)

    if using_copy_on_write:
        assert not np.shares_memory(
            get_array(ser, "a").left.values, get_array(view, "a").left.values
        )
        tm.assert_series_equal(view, ser_orig)
    else:
        assert np.shares_memory(
            get_array(ser, "a").left.values, get_array(view, "a").left.values
        )


def test_fillna_series_empty_arg(using_copy_on_write):
    ser = Series([1, np.nan, 2])
    ser_orig = ser.copy()
    result = ser.fillna({})

    if using_copy_on_write:
        assert np.shares_memory(get_array(ser), get_array(result))
    else:
        assert not np.shares_memory(get_array(ser), get_array(result))

    ser.iloc[0] = 100.5
    tm.assert_series_equal(ser_orig, result)


def test_fillna_series_empty_arg_inplace(using_copy_on_write):
    ser = Series([1, np.nan, 2])
    arr = get_array(ser)
    ser.fillna({}, inplace=True)

    assert np.shares_memory(get_array(ser), arr)
    if using_copy_on_write:
        assert ser._mgr._has_no_reference(0)


def test_fillna_ea_noop_shares_memory(
    using_copy_on_write, any_numeric_ea_and_arrow_dtype
):
    df = DataFrame({"a": [1, NA, 3], "b": 1}, dtype=any_numeric_ea_and_arrow_dtype)
    df_orig = df.copy()
    df2 = df.fillna(100)

    assert not np.shares_memory(get_array(df, "a"), get_array(df2, "a"))

    if using_copy_on_write:
        assert np.shares_memory(get_array(df, "b"), get_array(df2, "b"))
        assert not df2._mgr._has_no_reference(1)
    elif isinstance(df.dtypes.iloc[0], ArrowDtype):
        # arrow is immutable, so no-ops do not need to copy underlying array
        assert np.shares_memory(get_array(df, "b"), get_array(df2, "b"))
    else:
        assert not np.shares_memory(get_array(df, "b"), get_array(df2, "b"))

    tm.assert_frame_equal(df_orig, df)

    df2.iloc[0, 1] = 100
    if using_copy_on_write:
        assert not np.shares_memory(get_array(df, "b"), get_array(df2, "b"))
        assert df2._mgr._has_no_reference(1)
        assert df._mgr._has_no_reference(1)
    tm.assert_frame_equal(df_orig, df)


def test_fillna_inplace_ea_noop_shares_memory(
    using_copy_on_write, warn_copy_on_write, any_numeric_ea_and_arrow_dtype
):
    df = DataFrame({"a": [1, NA, 3], "b": 1}, dtype=any_numeric_ea_and_arrow_dtype)
    df_orig = df.copy()
    view = df[:]
    with tm.assert_cow_warning(warn_copy_on_write):
        df.fillna(100, inplace=True)

    if isinstance(df["a"].dtype, ArrowDtype) or using_copy_on_write:
        assert not np.shares_memory(get_array(df, "a"), get_array(view, "a"))
    else:
        # MaskedArray can actually respect inplace=True
        assert np.shares_memory(get_array(df, "a"), get_array(view, "a"))

    assert np.shares_memory(get_array(df, "b"), get_array(view, "b"))
    if using_copy_on_write:
        assert not df._mgr._has_no_reference(1)
        assert not view._mgr._has_no_reference(1)

    with tm.assert_cow_warning(
        warn_copy_on_write and "pyarrow" not in any_numeric_ea_and_arrow_dtype
    ):
        df.iloc[0, 1] = 100
    if isinstance(df["a"].dtype, ArrowDtype) or using_copy_on_write:
        tm.assert_frame_equal(df_orig, view)
    else:
        # we actually have a view
        tm.assert_frame_equal(df, view)


def test_fillna_chained_assignment(using_copy_on_write):
    df = DataFrame({"a": [1, np.nan, 2], "b": 1})
    df_orig = df.copy()
    if using_copy_on_write:
        with tm.raises_chained_assignment_error():
            df["a"].fillna(100, inplace=True)
        tm.assert_frame_equal(df, df_orig)

        with tm.raises_chained_assignment_error():
            df[["a"]].fillna(100, inplace=True)
        tm.assert_frame_equal(df, df_orig)
    else:
        with tm.assert_produces_warning(None):
            with option_context("mode.chained_assignment", None):
                df[["a"]].fillna(100, inplace=True)

        with tm.assert_produces_warning(None):
            with option_context("mode.chained_assignment", None):
                df[df.a > 5].fillna(100, inplace=True)

        with tm.assert_produces_warning(FutureWarning, match="inplace method"):
            df["a"].fillna(100, inplace=True)


@pytest.mark.parametrize("func", ["interpolate", "ffill", "bfill"])
def test_interpolate_chained_assignment(using_copy_on_write, func):
    df = DataFrame({"a": [1, np.nan, 2], "b": 1})
    df_orig = df.copy()
    if using_copy_on_write:
        with tm.raises_chained_assignment_error():
            getattr(df["a"], func)(inplace=True)
        tm.assert_frame_equal(df, df_orig)

        with tm.raises_chained_assignment_error():
            getattr(df[["a"]], func)(inplace=True)
        tm.assert_frame_equal(df, df_orig)
    else:
        with tm.assert_produces_warning(FutureWarning, match="inplace method"):
            getattr(df["a"], func)(inplace=True)

        with tm.assert_produces_warning(None):
            with option_context("mode.chained_assignment", None):
                getattr(df[["a"]], func)(inplace=True)

        with tm.assert_produces_warning(None):
            with option_context("mode.chained_assignment", None):
                getattr(df[df["a"] > 1], func)(inplace=True)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_core\tests\test_cpu_features.py ===
import os
import pathlib
import platform
import re
import subprocess
import sys

import pytest
from numpy._core._multiarray_umath import (
    __cpu_baseline__,
    __cpu_dispatch__,
    __cpu_features__,
)


def assert_features_equal(actual, desired, fname):
    __tracebackhide__ = True  # Hide traceback for py.test
    actual, desired = str(actual), str(desired)
    if actual == desired:
        return
    detected = str(__cpu_features__).replace("'", "")
    try:
        with open("/proc/cpuinfo") as fd:
            cpuinfo = fd.read(2048)
    except Exception as err:
        cpuinfo = str(err)

    try:
        import subprocess
        auxv = subprocess.check_output(['/bin/true'], env={"LD_SHOW_AUXV": "1"})
        auxv = auxv.decode()
    except Exception as err:
        auxv = str(err)

    import textwrap
    error_report = textwrap.indent(
f"""
###########################################
### Extra debugging information
###########################################
-------------------------------------------
--- NumPy Detections
-------------------------------------------
{detected}
-------------------------------------------
--- SYS / CPUINFO
-------------------------------------------
{cpuinfo}....
-------------------------------------------
--- SYS / AUXV
-------------------------------------------
{auxv}
""", prefix='\r')

    raise AssertionError((
        "Failure Detection\n"
        " NAME: '%s'\n"
        " ACTUAL: %s\n"
        " DESIRED: %s\n"
        "%s"
    ) % (fname, actual, desired, error_report))

def _text_to_list(txt):
    out = txt.strip("][\n").replace("'", "").split(', ')
    return None if out[0] == "" else out

class AbstractTest:
    features = []
    features_groups = {}
    features_map = {}
    features_flags = set()

    def load_flags(self):
        # a hook
        pass

    def test_features(self):
        self.load_flags()
        for gname, features in self.features_groups.items():
            test_features = [self.cpu_have(f) for f in features]
            assert_features_equal(__cpu_features__.get(gname), all(test_features), gname)

        for feature_name in self.features:
            cpu_have = self.cpu_have(feature_name)
            npy_have = __cpu_features__.get(feature_name)
            assert_features_equal(npy_have, cpu_have, feature_name)

    def cpu_have(self, feature_name):
        map_names = self.features_map.get(feature_name, feature_name)
        if isinstance(map_names, str):
            return map_names in self.features_flags
        return any(f in self.features_flags for f in map_names)

    def load_flags_cpuinfo(self, magic_key):
        self.features_flags = self.get_cpuinfo_item(magic_key)

    def get_cpuinfo_item(self, magic_key):
        values = set()
        with open('/proc/cpuinfo') as fd:
            for line in fd:
                if not line.startswith(magic_key):
                    continue
                flags_value = [s.strip() for s in line.split(':', 1)]
                if len(flags_value) == 2:
                    values = values.union(flags_value[1].upper().split())
        return values

    def load_flags_auxv(self):
        auxv = subprocess.check_output(['/bin/true'], env={"LD_SHOW_AUXV": "1"})
        for at in auxv.split(b'\n'):
            if not at.startswith(b"AT_HWCAP"):
                continue
            hwcap_value = [s.strip() for s in at.split(b':', 1)]
            if len(hwcap_value) == 2:
                self.features_flags = self.features_flags.union(
                    hwcap_value[1].upper().decode().split()
                )

@pytest.mark.skipif(
    sys.platform == 'emscripten',
    reason=(
        "The subprocess module is not available on WASM platforms and"
        " therefore this test class cannot be properly executed."
    ),
)
class TestEnvPrivation:
    cwd = pathlib.Path(__file__).parent.resolve()
    env = os.environ.copy()
    _enable = os.environ.pop('NPY_ENABLE_CPU_FEATURES', None)
    _disable = os.environ.pop('NPY_DISABLE_CPU_FEATURES', None)
    SUBPROCESS_ARGS = {"cwd": cwd, "capture_output": True, "text": True, "check": True}
    unavailable_feats = [
        feat for feat in __cpu_dispatch__ if not __cpu_features__[feat]
    ]
    UNAVAILABLE_FEAT = (
        None if len(unavailable_feats) == 0
        else unavailable_feats[0]
    )
    BASELINE_FEAT = None if len(__cpu_baseline__) == 0 else __cpu_baseline__[0]
    SCRIPT = """
def main():
    from numpy._core._multiarray_umath import (
        __cpu_features__,
        __cpu_dispatch__
    )

    detected = [feat for feat in __cpu_dispatch__ if __cpu_features__[feat]]
    print(detected)

if __name__ == "__main__":
    main()
    """

    @pytest.fixture(autouse=True)
    def setup_class(self, tmp_path_factory):
        file = tmp_path_factory.mktemp("runtime_test_script")
        file /= "_runtime_detect.py"
        file.write_text(self.SCRIPT)
        self.file = file

    def _run(self):
        return subprocess.run(
            [sys.executable, self.file],
            env=self.env,
            **self.SUBPROCESS_ARGS,
            )

    # Helper function mimicking pytest.raises for subprocess call
    def _expect_error(
        self,
        msg,
        err_type,
        no_error_msg="Failed to generate error"
    ):
        try:
            self._run()
        except subprocess.CalledProcessError as e:
            assertion_message = f"Expected: {msg}\nGot: {e.stderr}"
            assert re.search(msg, e.stderr), assertion_message

            assertion_message = (
                f"Expected error of type: {err_type}; see full "
                f"error:\n{e.stderr}"
            )
            assert re.search(err_type, e.stderr), assertion_message
        else:
            assert False, no_error_msg

    def setup_method(self):
        """Ensure that the environment is reset"""
        self.env = os.environ.copy()

    def test_runtime_feature_selection(self):
        """
        Ensure that when selecting `NPY_ENABLE_CPU_FEATURES`, only the
        features exactly specified are dispatched.
        """

        # Capture runtime-enabled features
        out = self._run()
        non_baseline_features = _text_to_list(out.stdout)

        if non_baseline_features is None:
            pytest.skip(
                "No dispatchable features outside of baseline detected."
            )
        feature = non_baseline_features[0]

        # Capture runtime-enabled features when `NPY_ENABLE_CPU_FEATURES` is
        # specified
        self.env['NPY_ENABLE_CPU_FEATURES'] = feature
        out = self._run()
        enabled_features = _text_to_list(out.stdout)

        # Ensure that only one feature is enabled, and it is exactly the one
        # specified by `NPY_ENABLE_CPU_FEATURES`
        assert set(enabled_features) == {feature}

        if len(non_baseline_features) < 2:
            pytest.skip("Only one non-baseline feature detected.")
        # Capture runtime-enabled features when `NPY_ENABLE_CPU_FEATURES` is
        # specified
        self.env['NPY_ENABLE_CPU_FEATURES'] = ",".join(non_baseline_features)
        out = self._run()
        enabled_features = _text_to_list(out.stdout)

        # Ensure that both features are enabled, and they are exactly the ones
        # specified by `NPY_ENABLE_CPU_FEATURES`
        assert set(enabled_features) == set(non_baseline_features)

    @pytest.mark.parametrize("enabled, disabled",
    [
        ("feature", "feature"),
        ("feature", "same"),
    ])
    def test_both_enable_disable_set(self, enabled, disabled):
        """
        Ensure that when both environment variables are set then an
        ImportError is thrown
        """
        self.env['NPY_ENABLE_CPU_FEATURES'] = enabled
        self.env['NPY_DISABLE_CPU_FEATURES'] = disabled
        msg = "Both NPY_DISABLE_CPU_FEATURES and NPY_ENABLE_CPU_FEATURES"
        err_type = "ImportError"
        self._expect_error(msg, err_type)

    @pytest.mark.skipif(
        not __cpu_dispatch__,
        reason=(
            "NPY_*_CPU_FEATURES only parsed if "
            "`__cpu_dispatch__` is non-empty"
        )
    )
    @pytest.mark.parametrize("action", ["ENABLE", "DISABLE"])
    def test_variable_too_long(self, action):
        """
        Test that an error is thrown if the environment variables are too long
        to be processed. Current limit is 1024, but this may change later.
        """
        MAX_VAR_LENGTH = 1024
        # Actual length is MAX_VAR_LENGTH + 1 due to null-termination
        self.env[f'NPY_{action}_CPU_FEATURES'] = "t" * MAX_VAR_LENGTH
        msg = (
            f"Length of environment variable 'NPY_{action}_CPU_FEATURES' is "
            f"{MAX_VAR_LENGTH + 1}, only {MAX_VAR_LENGTH} accepted"
        )
        err_type = "RuntimeError"
        self._expect_error(msg, err_type)

    @pytest.mark.skipif(
        not __cpu_dispatch__,
        reason=(
            "NPY_*_CPU_FEATURES only parsed if "
            "`__cpu_dispatch__` is non-empty"
        )
    )
    def test_impossible_feature_disable(self):
        """
        Test that a RuntimeError is thrown if an impossible feature-disabling
        request is made. This includes disabling a baseline feature.
        """

        if self.BASELINE_FEAT is None:
            pytest.skip("There are no unavailable features to test with")
        bad_feature = self.BASELINE_FEAT
        self.env['NPY_DISABLE_CPU_FEATURES'] = bad_feature
        msg = (
            f"You cannot disable CPU feature '{bad_feature}', since it is "
            "part of the baseline optimizations"
        )
        err_type = "RuntimeError"
        self._expect_error(msg, err_type)

    def test_impossible_feature_enable(self):
        """
        Test that a RuntimeError is thrown if an impossible feature-enabling
        request is made. This includes enabling a feature not supported by the
        machine, or disabling a baseline optimization.
        """

        if self.UNAVAILABLE_FEAT is None:
            pytest.skip("There are no unavailable features to test with")
        bad_feature = self.UNAVAILABLE_FEAT
        self.env['NPY_ENABLE_CPU_FEATURES'] = bad_feature
        msg = (
            f"You cannot enable CPU features \\({bad_feature}\\), since "
            "they are not supported by your machine."
        )
        err_type = "RuntimeError"
        self._expect_error(msg, err_type)

        # Ensure that it fails even when providing garbage in addition
        feats = f"{bad_feature}, Foobar"
        self.env['NPY_ENABLE_CPU_FEATURES'] = feats
        msg = (
            f"You cannot enable CPU features \\({bad_feature}\\), since they "
            "are not supported by your machine."
        )
        self._expect_error(msg, err_type)

        if self.BASELINE_FEAT is not None:
            # Ensure that only the bad feature gets reported
            feats = f"{bad_feature}, {self.BASELINE_FEAT}"
            self.env['NPY_ENABLE_CPU_FEATURES'] = feats
            msg = (
                f"You cannot enable CPU features \\({bad_feature}\\), since "
                "they are not supported by your machine."
            )
            self._expect_error(msg, err_type)


is_linux = sys.platform.startswith('linux')
is_cygwin = sys.platform.startswith('cygwin')
machine = platform.machine()
is_x86 = re.match(r"^(amd64|x86|i386|i686)", machine, re.IGNORECASE)
@pytest.mark.skipif(
    not (is_linux or is_cygwin) or not is_x86, reason="Only for Linux and x86"
)
class Test_X86_Features(AbstractTest):
    features = [
        "MMX", "SSE", "SSE2", "SSE3", "SSSE3", "SSE41", "POPCNT", "SSE42",
        "AVX", "F16C", "XOP", "FMA4", "FMA3", "AVX2", "AVX512F", "AVX512CD",
        "AVX512ER", "AVX512PF", "AVX5124FMAPS", "AVX5124VNNIW", "AVX512VPOPCNTDQ",
        "AVX512VL", "AVX512BW", "AVX512DQ", "AVX512VNNI", "AVX512IFMA",
        "AVX512VBMI", "AVX512VBMI2", "AVX512BITALG", "AVX512FP16",
    ]
    features_groups = {
        "AVX512_KNL": ["AVX512F", "AVX512CD", "AVX512ER", "AVX512PF"],
        "AVX512_KNM": ["AVX512F", "AVX512CD", "AVX512ER", "AVX512PF", "AVX5124FMAPS",
                      "AVX5124VNNIW", "AVX512VPOPCNTDQ"],
        "AVX512_SKX": ["AVX512F", "AVX512CD", "AVX512BW", "AVX512DQ", "AVX512VL"],
        "AVX512_CLX": ["AVX512F", "AVX512CD", "AVX512BW", "AVX512DQ", "AVX512VL", "AVX512VNNI"],
        "AVX512_CNL": ["AVX512F", "AVX512CD", "AVX512BW", "AVX512DQ", "AVX512VL", "AVX512IFMA",
                      "AVX512VBMI"],
        "AVX512_ICL": ["AVX512F", "AVX512CD", "AVX512BW", "AVX512DQ", "AVX512VL", "AVX512IFMA",
                      "AVX512VBMI", "AVX512VNNI", "AVX512VBMI2", "AVX512BITALG", "AVX512VPOPCNTDQ"],
        "AVX512_SPR": ["AVX512F", "AVX512CD", "AVX512BW", "AVX512DQ",
                      "AVX512VL", "AVX512IFMA", "AVX512VBMI", "AVX512VNNI",
                      "AVX512VBMI2", "AVX512BITALG", "AVX512VPOPCNTDQ",
                      "AVX512FP16"],
    }
    features_map = {
        "SSE3": "PNI", "SSE41": "SSE4_1", "SSE42": "SSE4_2", "FMA3": "FMA",
        "AVX512VNNI": "AVX512_VNNI", "AVX512BITALG": "AVX512_BITALG",
        "AVX512VBMI2": "AVX512_VBMI2", "AVX5124FMAPS": "AVX512_4FMAPS",
        "AVX5124VNNIW": "AVX512_4VNNIW", "AVX512VPOPCNTDQ": "AVX512_VPOPCNTDQ",
        "AVX512FP16": "AVX512_FP16",
    }

    def load_flags(self):
        self.load_flags_cpuinfo("flags")


is_power = re.match(r"^(powerpc|ppc)64", machine, re.IGNORECASE)
@pytest.mark.skipif(not is_linux or not is_power, reason="Only for Linux and Power")
class Test_POWER_Features(AbstractTest):
    features = ["VSX", "VSX2", "VSX3", "VSX4"]
    features_map = {"VSX2": "ARCH_2_07", "VSX3": "ARCH_3_00", "VSX4": "ARCH_3_1"}

    def load_flags(self):
        self.load_flags_auxv()


is_zarch = re.match(r"^(s390x)", machine, re.IGNORECASE)
@pytest.mark.skipif(not is_linux or not is_zarch,
                    reason="Only for Linux and IBM Z")
class Test_ZARCH_Features(AbstractTest):
    features = ["VX", "VXE", "VXE2"]

    def load_flags(self):
        self.load_flags_auxv()


is_arm = re.match(r"^(arm|aarch64)", machine, re.IGNORECASE)
@pytest.mark.skipif(not is_linux or not is_arm, reason="Only for Linux and ARM")
class Test_ARM_Features(AbstractTest):
    features = [
        "SVE", "NEON", "ASIMD", "FPHP", "ASIMDHP", "ASIMDDP", "ASIMDFHM"
    ]
    features_groups = {
        "NEON_FP16":  ["NEON", "HALF"],
        "NEON_VFPV4": ["NEON", "VFPV4"],
    }

    def load_flags(self):
        self.load_flags_cpuinfo("Features")
        arch = self.get_cpuinfo_item("CPU architecture")
        # in case of mounting virtual filesystem of aarch64 kernel without linux32
        is_rootfs_v8 = (
            not re.match(r"^armv[0-9]+l$", machine) and
            (int('0' + next(iter(arch))) > 7 if arch else 0)
        )
        if re.match(r"^(aarch64|AARCH64)", machine) or is_rootfs_v8:
            self.features_map = {
                "NEON": "ASIMD", "HALF": "ASIMD", "VFPV4": "ASIMD"
            }
        else:
            self.features_map = {
                # ELF auxiliary vector and /proc/cpuinfo on Linux kernel(armv8 aarch32)
                # doesn't provide information about ASIMD, so we assume that ASIMD is supported
                # if the kernel reports any one of the following ARM8 features.
                "ASIMD": ("AES", "SHA1", "SHA2", "PMULL", "CRC32")
            }


is_loongarch = re.match(r"^(loongarch)", machine, re.IGNORECASE)
@pytest.mark.skipif(not is_linux or not is_loongarch, reason="Only for Linux and LoongArch")
class Test_LOONGARCH_Features(AbstractTest):
    features = ["LSX"]

    def load_flags(self):
        self.load_flags_cpuinfo("Features")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\indexes\test_index_new.py ===
"""
Tests for the Index constructor conducting inference.
"""
from datetime import (
    datetime,
    timedelta,
    timezone,
)
from decimal import Decimal

import numpy as np
import pytest

from pandas._libs.tslibs.timezones import maybe_get_tz

from pandas import (
    NA,
    Categorical,
    CategoricalIndex,
    DatetimeIndex,
    Index,
    IntervalIndex,
    MultiIndex,
    NaT,
    PeriodIndex,
    Series,
    TimedeltaIndex,
    Timestamp,
    array,
    date_range,
    period_range,
    timedelta_range,
)
import pandas._testing as tm


class TestIndexConstructorInference:
    def test_object_all_bools(self):
        # GH#49594 match Series behavior on ndarray[object] of all bools
        arr = np.array([True, False], dtype=object)
        res = Index(arr)
        assert res.dtype == object

        # since the point is matching Series behavior, let's double check
        assert Series(arr).dtype == object

    def test_object_all_complex(self):
        # GH#49594 match Series behavior on ndarray[object] of all complex
        arr = np.array([complex(1), complex(2)], dtype=object)
        res = Index(arr)
        assert res.dtype == object

        # since the point is matching Series behavior, let's double check
        assert Series(arr).dtype == object

    @pytest.mark.parametrize("val", [NaT, None, np.nan, float("nan")])
    def test_infer_nat(self, val):
        # GH#49340 all NaT/None/nan and at least 1 NaT -> datetime64[ns],
        #  matching Series behavior
        values = [NaT, val]

        idx = Index(values)
        assert idx.dtype == "datetime64[ns]" and idx.isna().all()

        idx = Index(values[::-1])
        assert idx.dtype == "datetime64[ns]" and idx.isna().all()

        idx = Index(np.array(values, dtype=object))
        assert idx.dtype == "datetime64[ns]" and idx.isna().all()

        idx = Index(np.array(values, dtype=object)[::-1])
        assert idx.dtype == "datetime64[ns]" and idx.isna().all()

    @pytest.mark.parametrize("na_value", [None, np.nan])
    @pytest.mark.parametrize("vtype", [list, tuple, iter])
    def test_construction_list_tuples_nan(self, na_value, vtype):
        # GH#18505 : valid tuples containing NaN
        values = [(1, "two"), (3.0, na_value)]
        result = Index(vtype(values))
        expected = MultiIndex.from_tuples(values)
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "dtype",
        [int, "int64", "int32", "int16", "int8", "uint64", "uint32", "uint16", "uint8"],
    )
    def test_constructor_int_dtype_float(self, dtype):
        # GH#18400
        expected = Index([0, 1, 2, 3], dtype=dtype)
        result = Index([0.0, 1.0, 2.0, 3.0], dtype=dtype)
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize("cast_index", [True, False])
    @pytest.mark.parametrize(
        "vals", [[True, False, True], np.array([True, False, True], dtype=bool)]
    )
    def test_constructor_dtypes_to_object(self, cast_index, vals):
        if cast_index:
            index = Index(vals, dtype=bool)
        else:
            index = Index(vals)

        assert type(index) is Index
        assert index.dtype == bool

    def test_constructor_categorical_to_object(self):
        # GH#32167 Categorical data and dtype=object should return object-dtype
        ci = CategoricalIndex(range(5))
        result = Index(ci, dtype=object)
        assert not isinstance(result, CategoricalIndex)

    def test_constructor_infer_periodindex(self):
        xp = period_range("2012-1-1", freq="M", periods=3)
        rs = Index(xp)
        tm.assert_index_equal(rs, xp)
        assert isinstance(rs, PeriodIndex)

    def test_from_list_of_periods(self):
        rng = period_range("1/1/2000", periods=20, freq="D")
        periods = list(rng)

        result = Index(periods)
        assert isinstance(result, PeriodIndex)

    @pytest.mark.parametrize("pos", [0, 1])
    @pytest.mark.parametrize(
        "klass,dtype,ctor",
        [
            (DatetimeIndex, "datetime64[ns]", np.datetime64("nat")),
            (TimedeltaIndex, "timedelta64[ns]", np.timedelta64("nat")),
        ],
    )
    def test_constructor_infer_nat_dt_like(
        self, pos, klass, dtype, ctor, nulls_fixture, request
    ):
        if isinstance(nulls_fixture, Decimal):
            # We dont cast these to datetime64/timedelta64
            pytest.skip(
                f"We don't cast {type(nulls_fixture).__name__} to "
                "datetime64/timedelta64"
            )

        expected = klass([NaT, NaT])
        assert expected.dtype == dtype
        data = [ctor]
        data.insert(pos, nulls_fixture)

        warn = None
        if nulls_fixture is NA:
            expected = Index([NA, NaT])
            mark = pytest.mark.xfail(reason="Broken with np.NaT ctor; see GH 31884")
            request.applymarker(mark)
            # GH#35942 numpy will emit a DeprecationWarning within the
            #  assert_index_equal calls.  Since we can't do anything
            #  about it until GH#31884 is fixed, we suppress that warning.
            warn = DeprecationWarning

        result = Index(data)

        with tm.assert_produces_warning(warn):
            tm.assert_index_equal(result, expected)

        result = Index(np.array(data, dtype=object))

        with tm.assert_produces_warning(warn):
            tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize("swap_objs", [True, False])
    def test_constructor_mixed_nat_objs_infers_object(self, swap_objs):
        # mixed np.datetime64/timedelta64 nat results in object
        data = [np.datetime64("nat"), np.timedelta64("nat")]
        if swap_objs:
            data = data[::-1]

        expected = Index(data, dtype=object)
        tm.assert_index_equal(Index(data), expected)
        tm.assert_index_equal(Index(np.array(data, dtype=object)), expected)

    @pytest.mark.parametrize("swap_objs", [True, False])
    def test_constructor_datetime_and_datetime64(self, swap_objs):
        data = [Timestamp(2021, 6, 8, 9, 42), np.datetime64("now")]
        if swap_objs:
            data = data[::-1]
        expected = DatetimeIndex(data)

        tm.assert_index_equal(Index(data), expected)
        tm.assert_index_equal(Index(np.array(data, dtype=object)), expected)

    def test_constructor_datetimes_mixed_tzs(self):
        # https://github.com/pandas-dev/pandas/pull/55793/files#r1383719998
        tz = maybe_get_tz("US/Central")
        dt1 = datetime(2020, 1, 1, tzinfo=tz)
        dt2 = datetime(2020, 1, 1, tzinfo=timezone.utc)
        result = Index([dt1, dt2])
        expected = Index([dt1, dt2], dtype=object)
        tm.assert_index_equal(result, expected)


class TestDtypeEnforced:
    # check we don't silently ignore the dtype keyword

    def test_constructor_object_dtype_with_ea_data(self, any_numeric_ea_dtype):
        # GH#45206
        arr = array([0], dtype=any_numeric_ea_dtype)

        idx = Index(arr, dtype=object)
        assert idx.dtype == object

    @pytest.mark.parametrize("dtype", [object, "float64", "uint64", "category"])
    def test_constructor_range_values_mismatched_dtype(self, dtype):
        rng = Index(range(5))

        result = Index(rng, dtype=dtype)
        assert result.dtype == dtype

        result = Index(range(5), dtype=dtype)
        assert result.dtype == dtype

    @pytest.mark.parametrize("dtype", [object, "float64", "uint64", "category"])
    def test_constructor_categorical_values_mismatched_non_ea_dtype(self, dtype):
        cat = Categorical([1, 2, 3])

        result = Index(cat, dtype=dtype)
        assert result.dtype == dtype

    def test_constructor_categorical_values_mismatched_dtype(self):
        dti = date_range("2016-01-01", periods=3)
        cat = Categorical(dti)
        result = Index(cat, dti.dtype)
        tm.assert_index_equal(result, dti)

        dti2 = dti.tz_localize("Asia/Tokyo")
        cat2 = Categorical(dti2)
        result = Index(cat2, dti2.dtype)
        tm.assert_index_equal(result, dti2)

        ii = IntervalIndex.from_breaks(range(5))
        cat3 = Categorical(ii)
        result = Index(cat3, dtype=ii.dtype)
        tm.assert_index_equal(result, ii)

    def test_constructor_ea_values_mismatched_categorical_dtype(self):
        dti = date_range("2016-01-01", periods=3)
        result = Index(dti, dtype="category")
        expected = CategoricalIndex(dti)
        tm.assert_index_equal(result, expected)

        dti2 = date_range("2016-01-01", periods=3, tz="US/Pacific")
        result = Index(dti2, dtype="category")
        expected = CategoricalIndex(dti2)
        tm.assert_index_equal(result, expected)

    def test_constructor_period_values_mismatched_dtype(self):
        pi = period_range("2016-01-01", periods=3, freq="D")
        result = Index(pi, dtype="category")
        expected = CategoricalIndex(pi)
        tm.assert_index_equal(result, expected)

    def test_constructor_timedelta64_values_mismatched_dtype(self):
        # check we don't silently ignore the dtype keyword
        tdi = timedelta_range("4 Days", periods=5)
        result = Index(tdi, dtype="category")
        expected = CategoricalIndex(tdi)
        tm.assert_index_equal(result, expected)

    def test_constructor_interval_values_mismatched_dtype(self):
        dti = date_range("2016-01-01", periods=3)
        ii = IntervalIndex.from_breaks(dti)
        result = Index(ii, dtype="category")
        expected = CategoricalIndex(ii)
        tm.assert_index_equal(result, expected)

    def test_constructor_datetime64_values_mismatched_period_dtype(self):
        dti = date_range("2016-01-01", periods=3)
        result = Index(dti, dtype="Period[D]")
        expected = dti.to_period("D")
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize("dtype", ["int64", "uint64"])
    def test_constructor_int_dtype_nan_raises(self, dtype):
        # see GH#15187
        data = [np.nan]
        msg = "cannot convert"
        with pytest.raises(ValueError, match=msg):
            Index(data, dtype=dtype)

    @pytest.mark.parametrize(
        "vals",
        [
            [1, 2, 3],
            np.array([1, 2, 3]),
            np.array([1, 2, 3], dtype=int),
            # below should coerce
            [1.0, 2.0, 3.0],
            np.array([1.0, 2.0, 3.0], dtype=float),
        ],
    )
    def test_constructor_dtypes_to_int(self, vals, any_int_numpy_dtype):
        dtype = any_int_numpy_dtype
        index = Index(vals, dtype=dtype)
        assert index.dtype == dtype

    @pytest.mark.parametrize(
        "vals",
        [
            [1, 2, 3],
            [1.0, 2.0, 3.0],
            np.array([1.0, 2.0, 3.0]),
            np.array([1, 2, 3], dtype=int),
            np.array([1.0, 2.0, 3.0], dtype=float),
        ],
    )
    def test_constructor_dtypes_to_float(self, vals, float_numpy_dtype):
        dtype = float_numpy_dtype
        index = Index(vals, dtype=dtype)
        assert index.dtype == dtype

    @pytest.mark.parametrize(
        "vals",
        [
            [1, 2, 3],
            np.array([1, 2, 3], dtype=int),
            np.array(["2011-01-01", "2011-01-02"], dtype="datetime64[ns]"),
            [datetime(2011, 1, 1), datetime(2011, 1, 2)],
        ],
    )
    def test_constructor_dtypes_to_categorical(self, vals):
        index = Index(vals, dtype="category")
        assert isinstance(index, CategoricalIndex)

    @pytest.mark.parametrize("cast_index", [True, False])
    @pytest.mark.parametrize(
        "vals",
        [
            Index(np.array([np.datetime64("2011-01-01"), np.datetime64("2011-01-02")])),
            Index([datetime(2011, 1, 1), datetime(2011, 1, 2)]),
        ],
    )
    def test_constructor_dtypes_to_datetime(self, cast_index, vals):
        if cast_index:
            index = Index(vals, dtype=object)
            assert isinstance(index, Index)
            assert index.dtype == object
        else:
            index = Index(vals)
            assert isinstance(index, DatetimeIndex)

    @pytest.mark.parametrize("cast_index", [True, False])
    @pytest.mark.parametrize(
        "vals",
        [
            np.array([np.timedelta64(1, "D"), np.timedelta64(1, "D")]),
            [timedelta(1), timedelta(1)],
        ],
    )
    def test_constructor_dtypes_to_timedelta(self, cast_index, vals):
        if cast_index:
            index = Index(vals, dtype=object)
            assert isinstance(index, Index)
            assert index.dtype == object
        else:
            index = Index(vals)
            assert isinstance(index, TimedeltaIndex)

    def test_pass_timedeltaindex_to_index(self):
        rng = timedelta_range("1 days", "10 days")
        idx = Index(rng, dtype=object)

        expected = Index(rng.to_pytimedelta(), dtype=object)

        tm.assert_numpy_array_equal(idx.values, expected.values)

    def test_pass_datetimeindex_to_index(self):
        # GH#1396
        rng = date_range("1/1/2000", "3/1/2000")
        idx = Index(rng, dtype=object)

        expected = Index(rng.to_pydatetime(), dtype=object)

        tm.assert_numpy_array_equal(idx.values, expected.values)


class TestIndexConstructorUnwrapping:
    # Test passing different arraylike values to pd.Index

    @pytest.mark.parametrize("klass", [Index, DatetimeIndex])
    def test_constructor_from_series_dt64(self, klass):
        stamps = [Timestamp("20110101"), Timestamp("20120101"), Timestamp("20130101")]
        expected = DatetimeIndex(stamps)
        ser = Series(stamps)
        result = klass(ser)
        tm.assert_index_equal(result, expected)

    def test_constructor_no_pandas_array(self):
        ser = Series([1, 2, 3])
        result = Index(ser.array)
        expected = Index([1, 2, 3])
        tm.assert_index_equal(result, expected)

    @pytest.mark.parametrize(
        "array",
        [
            np.arange(5),
            np.array(["a", "b", "c"]),
            date_range("2000-01-01", periods=3).values,
        ],
    )
    def test_constructor_ndarray_like(self, array):
        # GH#5460#issuecomment-44474502
        # it should be possible to convert any object that satisfies the numpy
        # ndarray interface directly into an Index
        class ArrayLike:
            def __init__(self, array) -> None:
                self.array = array

            def __array__(self, dtype=None, copy=None) -> np.ndarray:
                return self.array

        expected = Index(array)
        result = Index(ArrayLike(array))
        tm.assert_index_equal(result, expected)


class TestIndexConstructionErrors:
    def test_constructor_overflow_int64(self):
        # see GH#15832
        msg = (
            "The elements provided in the data cannot "
            "all be casted to the dtype int64"
        )
        with pytest.raises(OverflowError, match=msg):
            Index([np.iinfo(np.uint64).max - 1], dtype="int64")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\io\excel\test_openpyxl.py ===
import contextlib
from pathlib import Path
import re

import numpy as np
import pytest

from pandas.compat import is_platform_windows

import pandas as pd
from pandas import DataFrame
import pandas._testing as tm

from pandas.io.excel import (
    ExcelWriter,
    _OpenpyxlWriter,
)
from pandas.io.excel._openpyxl import OpenpyxlReader

openpyxl = pytest.importorskip("openpyxl")

if is_platform_windows():
    pytestmark = pytest.mark.single_cpu


@pytest.fixture
def ext():
    return ".xlsx"


def test_to_excel_styleconverter():
    from openpyxl import styles

    hstyle = {
        "font": {"color": "00FF0000", "bold": True},
        "borders": {"top": "thin", "right": "thin", "bottom": "thin", "left": "thin"},
        "alignment": {"horizontal": "center", "vertical": "top"},
        "fill": {"patternType": "solid", "fgColor": {"rgb": "006666FF", "tint": 0.3}},
        "number_format": {"format_code": "0.00"},
        "protection": {"locked": True, "hidden": False},
    }

    font_color = styles.Color("00FF0000")
    font = styles.Font(bold=True, color=font_color)
    side = styles.Side(style=styles.borders.BORDER_THIN)
    border = styles.Border(top=side, right=side, bottom=side, left=side)
    alignment = styles.Alignment(horizontal="center", vertical="top")
    fill_color = styles.Color(rgb="006666FF", tint=0.3)
    fill = styles.PatternFill(patternType="solid", fgColor=fill_color)

    number_format = "0.00"

    protection = styles.Protection(locked=True, hidden=False)

    kw = _OpenpyxlWriter._convert_to_style_kwargs(hstyle)
    assert kw["font"] == font
    assert kw["border"] == border
    assert kw["alignment"] == alignment
    assert kw["fill"] == fill
    assert kw["number_format"] == number_format
    assert kw["protection"] == protection


def test_write_cells_merge_styled(ext):
    from pandas.io.formats.excel import ExcelCell

    sheet_name = "merge_styled"

    sty_b1 = {"font": {"color": "00FF0000"}}
    sty_a2 = {"font": {"color": "0000FF00"}}

    initial_cells = [
        ExcelCell(col=1, row=0, val=42, style=sty_b1),
        ExcelCell(col=0, row=1, val=99, style=sty_a2),
    ]

    sty_merged = {"font": {"color": "000000FF", "bold": True}}
    sty_kwargs = _OpenpyxlWriter._convert_to_style_kwargs(sty_merged)
    openpyxl_sty_merged = sty_kwargs["font"]
    merge_cells = [
        ExcelCell(
            col=0, row=0, val="pandas", mergestart=1, mergeend=1, style=sty_merged
        )
    ]

    with tm.ensure_clean(ext) as path:
        with _OpenpyxlWriter(path) as writer:
            writer._write_cells(initial_cells, sheet_name=sheet_name)
            writer._write_cells(merge_cells, sheet_name=sheet_name)

            wks = writer.sheets[sheet_name]
        xcell_b1 = wks["B1"]
        xcell_a2 = wks["A2"]
        assert xcell_b1.font == openpyxl_sty_merged
        assert xcell_a2.font == openpyxl_sty_merged


@pytest.mark.parametrize("iso_dates", [True, False])
def test_engine_kwargs_write(ext, iso_dates):
    # GH 42286 GH 43445
    engine_kwargs = {"iso_dates": iso_dates}
    with tm.ensure_clean(ext) as f:
        with ExcelWriter(f, engine="openpyxl", engine_kwargs=engine_kwargs) as writer:
            assert writer.book.iso_dates == iso_dates
            # ExcelWriter won't allow us to close without writing something
            DataFrame().to_excel(writer)


def test_engine_kwargs_append_invalid(ext):
    # GH 43445
    # test whether an invalid engine kwargs actually raises
    with tm.ensure_clean(ext) as f:
        DataFrame(["hello", "world"]).to_excel(f)
        with pytest.raises(
            TypeError,
            match=re.escape(
                "load_workbook() got an unexpected keyword argument 'apple_banana'"
            ),
        ):
            with ExcelWriter(
                f, engine="openpyxl", mode="a", engine_kwargs={"apple_banana": "fruit"}
            ) as writer:
                # ExcelWriter needs us to write something to close properly
                DataFrame(["good"]).to_excel(writer, sheet_name="Sheet2")


@pytest.mark.parametrize("data_only, expected", [(True, 0), (False, "=1+1")])
def test_engine_kwargs_append_data_only(ext, data_only, expected):
    # GH 43445
    # tests whether the data_only engine_kwarg actually works well for
    # openpyxl's load_workbook
    with tm.ensure_clean(ext) as f:
        DataFrame(["=1+1"]).to_excel(f)
        with ExcelWriter(
            f, engine="openpyxl", mode="a", engine_kwargs={"data_only": data_only}
        ) as writer:
            assert writer.sheets["Sheet1"]["B2"].value == expected
            # ExcelWriter needs us to writer something to close properly?
            DataFrame().to_excel(writer, sheet_name="Sheet2")

        # ensure that data_only also works for reading
        #  and that formulas/values roundtrip
        assert (
            pd.read_excel(
                f,
                sheet_name="Sheet1",
                engine="openpyxl",
                engine_kwargs={"data_only": data_only},
            ).iloc[0, 1]
            == expected
        )


@pytest.mark.parametrize("kwarg_name", ["read_only", "data_only"])
@pytest.mark.parametrize("kwarg_value", [True, False])
def test_engine_kwargs_append_reader(datapath, ext, kwarg_name, kwarg_value):
    # GH 55027
    # test that `read_only` and `data_only` can be passed to
    #  `openpyxl.reader.excel.load_workbook` via `engine_kwargs`
    filename = datapath("io", "data", "excel", "test1" + ext)
    with contextlib.closing(
        OpenpyxlReader(filename, engine_kwargs={kwarg_name: kwarg_value})
    ) as reader:
        assert getattr(reader.book, kwarg_name) == kwarg_value


@pytest.mark.parametrize(
    "mode,expected", [("w", ["baz"]), ("a", ["foo", "bar", "baz"])]
)
def test_write_append_mode(ext, mode, expected):
    df = DataFrame([1], columns=["baz"])

    with tm.ensure_clean(ext) as f:
        wb = openpyxl.Workbook()
        wb.worksheets[0].title = "foo"
        wb.worksheets[0]["A1"].value = "foo"
        wb.create_sheet("bar")
        wb.worksheets[1]["A1"].value = "bar"
        wb.save(f)

        with ExcelWriter(f, engine="openpyxl", mode=mode) as writer:
            df.to_excel(writer, sheet_name="baz", index=False)

        with contextlib.closing(openpyxl.load_workbook(f)) as wb2:
            result = [sheet.title for sheet in wb2.worksheets]
            assert result == expected

            for index, cell_value in enumerate(expected):
                assert wb2.worksheets[index]["A1"].value == cell_value


@pytest.mark.parametrize(
    "if_sheet_exists,num_sheets,expected",
    [
        ("new", 2, ["apple", "banana"]),
        ("replace", 1, ["pear"]),
        ("overlay", 1, ["pear", "banana"]),
    ],
)
def test_if_sheet_exists_append_modes(ext, if_sheet_exists, num_sheets, expected):
    # GH 40230
    df1 = DataFrame({"fruit": ["apple", "banana"]})
    df2 = DataFrame({"fruit": ["pear"]})

    with tm.ensure_clean(ext) as f:
        df1.to_excel(f, engine="openpyxl", sheet_name="foo", index=False)
        with ExcelWriter(
            f, engine="openpyxl", mode="a", if_sheet_exists=if_sheet_exists
        ) as writer:
            df2.to_excel(writer, sheet_name="foo", index=False)

        with contextlib.closing(openpyxl.load_workbook(f)) as wb:
            assert len(wb.sheetnames) == num_sheets
            assert wb.sheetnames[0] == "foo"
            result = pd.read_excel(wb, "foo", engine="openpyxl")
            assert list(result["fruit"]) == expected
            if len(wb.sheetnames) == 2:
                result = pd.read_excel(wb, wb.sheetnames[1], engine="openpyxl")
                tm.assert_frame_equal(result, df2)


@pytest.mark.parametrize(
    "startrow, startcol, greeting, goodbye",
    [
        (0, 0, ["poop", "world"], ["goodbye", "people"]),
        (0, 1, ["hello", "world"], ["poop", "people"]),
        (1, 0, ["hello", "poop"], ["goodbye", "people"]),
        (1, 1, ["hello", "world"], ["goodbye", "poop"]),
    ],
)
def test_append_overlay_startrow_startcol(ext, startrow, startcol, greeting, goodbye):
    df1 = DataFrame({"greeting": ["hello", "world"], "goodbye": ["goodbye", "people"]})
    df2 = DataFrame(["poop"])

    with tm.ensure_clean(ext) as f:
        df1.to_excel(f, engine="openpyxl", sheet_name="poo", index=False)
        with ExcelWriter(
            f, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        ) as writer:
            # use startrow+1 because we don't have a header
            df2.to_excel(
                writer,
                index=False,
                header=False,
                startrow=startrow + 1,
                startcol=startcol,
                sheet_name="poo",
            )

        result = pd.read_excel(f, sheet_name="poo", engine="openpyxl")
        expected = DataFrame({"greeting": greeting, "goodbye": goodbye})
        tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "if_sheet_exists,msg",
    [
        (
            "invalid",
            "'invalid' is not valid for if_sheet_exists. Valid options "
            "are 'error', 'new', 'replace' and 'overlay'.",
        ),
        (
            "error",
            "Sheet 'foo' already exists and if_sheet_exists is set to 'error'.",
        ),
        (
            None,
            "Sheet 'foo' already exists and if_sheet_exists is set to 'error'.",
        ),
    ],
)
def test_if_sheet_exists_raises(ext, if_sheet_exists, msg):
    # GH 40230
    df = DataFrame({"fruit": ["pear"]})
    with tm.ensure_clean(ext) as f:
        with pytest.raises(ValueError, match=re.escape(msg)):
            df.to_excel(f, sheet_name="foo", engine="openpyxl")
            with ExcelWriter(
                f, engine="openpyxl", mode="a", if_sheet_exists=if_sheet_exists
            ) as writer:
                df.to_excel(writer, sheet_name="foo")


def test_to_excel_with_openpyxl_engine(ext):
    # GH 29854
    with tm.ensure_clean(ext) as filename:
        df1 = DataFrame({"A": np.linspace(1, 10, 10)})
        df2 = DataFrame({"B": np.linspace(1, 20, 10)})
        df = pd.concat([df1, df2], axis=1)
        styled = df.style.map(
            lambda val: f"color: {'red' if val < 0 else 'black'}"
        ).highlight_max()

        styled.to_excel(filename, engine="openpyxl")


@pytest.mark.parametrize("read_only", [True, False])
def test_read_workbook(datapath, ext, read_only):
    # GH 39528
    filename = datapath("io", "data", "excel", "test1" + ext)
    with contextlib.closing(
        openpyxl.load_workbook(filename, read_only=read_only)
    ) as wb:
        result = pd.read_excel(wb, engine="openpyxl")
    expected = pd.read_excel(filename)
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "header, expected_data",
    [
        (
            0,
            {
                "Title": [np.nan, "A", 1, 2, 3],
                "Unnamed: 1": [np.nan, "B", 4, 5, 6],
                "Unnamed: 2": [np.nan, "C", 7, 8, 9],
            },
        ),
        (2, {"A": [1, 2, 3], "B": [4, 5, 6], "C": [7, 8, 9]}),
    ],
)
@pytest.mark.parametrize(
    "filename", ["dimension_missing", "dimension_small", "dimension_large"]
)
# When read_only is None, use read_excel instead of a workbook
@pytest.mark.parametrize("read_only", [True, False, None])
def test_read_with_bad_dimension(
    datapath, ext, header, expected_data, filename, read_only
):
    # GH 38956, 39001 - no/incorrect dimension information
    path = datapath("io", "data", "excel", f"{filename}{ext}")
    if read_only is None:
        result = pd.read_excel(path, header=header)
    else:
        with contextlib.closing(
            openpyxl.load_workbook(path, read_only=read_only)
        ) as wb:
            result = pd.read_excel(wb, engine="openpyxl", header=header)
    expected = DataFrame(expected_data)
    tm.assert_frame_equal(result, expected)


def test_append_mode_file(ext):
    # GH 39576
    df = DataFrame()

    with tm.ensure_clean(ext) as f:
        df.to_excel(f, engine="openpyxl")

        with ExcelWriter(
            f, mode="a", engine="openpyxl", if_sheet_exists="new"
        ) as writer:
            df.to_excel(writer)

        # make sure that zip files are not concatenated by making sure that
        # "docProps/app.xml" only occurs twice in the file
        data = Path(f).read_bytes()
        first = data.find(b"docProps/app.xml")
        second = data.find(b"docProps/app.xml", first + 1)
        third = data.find(b"docProps/app.xml", second + 1)
        assert second != -1 and third == -1


# When read_only is None, use read_excel instead of a workbook
@pytest.mark.parametrize("read_only", [True, False, None])
def test_read_with_empty_trailing_rows(datapath, ext, read_only):
    # GH 39181
    path = datapath("io", "data", "excel", f"empty_trailing_rows{ext}")
    if read_only is None:
        result = pd.read_excel(path)
    else:
        with contextlib.closing(
            openpyxl.load_workbook(path, read_only=read_only)
        ) as wb:
            result = pd.read_excel(wb, engine="openpyxl")
    expected = DataFrame(
        {
            "Title": [np.nan, "A", 1, 2, 3],
            "Unnamed: 1": [np.nan, "B", 4, 5, 6],
            "Unnamed: 2": [np.nan, "C", 7, 8, 9],
        }
    )
    tm.assert_frame_equal(result, expected)


# When read_only is None, use read_excel instead of a workbook
@pytest.mark.parametrize("read_only", [True, False, None])
def test_read_empty_with_blank_row(datapath, ext, read_only):
    # GH 39547 - empty excel file with a row that has no data
    path = datapath("io", "data", "excel", f"empty_with_blank_row{ext}")
    if read_only is None:
        result = pd.read_excel(path)
    else:
        with contextlib.closing(
            openpyxl.load_workbook(path, read_only=read_only)
        ) as wb:
            result = pd.read_excel(wb, engine="openpyxl")
    expected = DataFrame()
    tm.assert_frame_equal(result, expected)


def test_book_and_sheets_consistent(ext):
    # GH#45687 - Ensure sheets is updated if user modifies book
    with tm.ensure_clean(ext) as f:
        with ExcelWriter(f, engine="openpyxl") as writer:
            assert writer.sheets == {}
            sheet = writer.book.create_sheet("test_name", 0)
            assert writer.sheets == {"test_name": sheet}


def test_ints_spelled_with_decimals(datapath, ext):
    # GH 46988 - openpyxl returns this sheet with floats
    path = datapath("io", "data", "excel", f"ints_spelled_with_decimals{ext}")
    result = pd.read_excel(path)
    expected = DataFrame(range(2, 12), columns=[1])
    tm.assert_frame_equal(result, expected)


def test_read_multiindex_header_no_index_names(datapath, ext):
    # GH#47487
    path = datapath("io", "data", "excel", f"multiindex_no_index_names{ext}")
    result = pd.read_excel(path, index_col=[0, 1, 2], header=[0, 1, 2])
    expected = DataFrame(
        [[np.nan, "x", "x", "x"], ["x", np.nan, np.nan, np.nan]],
        columns=pd.MultiIndex.from_tuples(
            [("X", "Y", "A1"), ("X", "Y", "A2"), ("XX", "YY", "B1"), ("XX", "YY", "B2")]
        ),
        index=pd.MultiIndex.from_tuples([("A", "AA", "AAA"), ("A", "BB", "BBB")]),
    )
    tm.assert_frame_equal(result, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\io\formats\test_to_excel.py ===
"""Tests formatting as writer-agnostic ExcelCells

ExcelFormatter is tested implicitly in pandas/tests/io/excel
"""
import string

import pytest

from pandas.errors import CSSWarning

import pandas._testing as tm

from pandas.io.formats.excel import (
    CssExcelCell,
    CSSToExcelConverter,
)


@pytest.mark.parametrize(
    "css,expected",
    [
        # FONT
        # - name
        ("font-family: foo,bar", {"font": {"name": "foo"}}),
        ('font-family: "foo bar",baz', {"font": {"name": "foo bar"}}),
        ("font-family: foo,\nbar", {"font": {"name": "foo"}}),
        ("font-family: foo, bar,    baz", {"font": {"name": "foo"}}),
        ("font-family: bar, foo", {"font": {"name": "bar"}}),
        ("font-family: 'foo bar', baz", {"font": {"name": "foo bar"}}),
        ("font-family: 'foo \\'bar', baz", {"font": {"name": "foo 'bar"}}),
        ('font-family: "foo \\"bar", baz', {"font": {"name": 'foo "bar'}}),
        ('font-family: "foo ,bar", baz', {"font": {"name": "foo ,bar"}}),
        # - family
        ("font-family: serif", {"font": {"name": "serif", "family": 1}}),
        ("font-family: Serif", {"font": {"name": "serif", "family": 1}}),
        ("font-family: roman, serif", {"font": {"name": "roman", "family": 1}}),
        ("font-family: roman, sans-serif", {"font": {"name": "roman", "family": 2}}),
        ("font-family: roman, sans serif", {"font": {"name": "roman"}}),
        ("font-family: roman, sansserif", {"font": {"name": "roman"}}),
        ("font-family: roman, cursive", {"font": {"name": "roman", "family": 4}}),
        ("font-family: roman, fantasy", {"font": {"name": "roman", "family": 5}}),
        # - size
        ("font-size: 1em", {"font": {"size": 12}}),
        ("font-size: xx-small", {"font": {"size": 6}}),
        ("font-size: x-small", {"font": {"size": 7.5}}),
        ("font-size: small", {"font": {"size": 9.6}}),
        ("font-size: medium", {"font": {"size": 12}}),
        ("font-size: large", {"font": {"size": 13.5}}),
        ("font-size: x-large", {"font": {"size": 18}}),
        ("font-size: xx-large", {"font": {"size": 24}}),
        ("font-size: 50%", {"font": {"size": 6}}),
        # - bold
        ("font-weight: 100", {"font": {"bold": False}}),
        ("font-weight: 200", {"font": {"bold": False}}),
        ("font-weight: 300", {"font": {"bold": False}}),
        ("font-weight: 400", {"font": {"bold": False}}),
        ("font-weight: normal", {"font": {"bold": False}}),
        ("font-weight: lighter", {"font": {"bold": False}}),
        ("font-weight: bold", {"font": {"bold": True}}),
        ("font-weight: bolder", {"font": {"bold": True}}),
        ("font-weight: 700", {"font": {"bold": True}}),
        ("font-weight: 800", {"font": {"bold": True}}),
        ("font-weight: 900", {"font": {"bold": True}}),
        # - italic
        ("font-style: italic", {"font": {"italic": True}}),
        ("font-style: oblique", {"font": {"italic": True}}),
        # - underline
        ("text-decoration: underline", {"font": {"underline": "single"}}),
        ("text-decoration: overline", {}),
        ("text-decoration: none", {}),
        # - strike
        ("text-decoration: line-through", {"font": {"strike": True}}),
        (
            "text-decoration: underline line-through",
            {"font": {"strike": True, "underline": "single"}},
        ),
        (
            "text-decoration: underline; text-decoration: line-through",
            {"font": {"strike": True}},
        ),
        # - color
        ("color: red", {"font": {"color": "FF0000"}}),
        ("color: #ff0000", {"font": {"color": "FF0000"}}),
        ("color: #f0a", {"font": {"color": "FF00AA"}}),
        # - shadow
        ("text-shadow: none", {"font": {"shadow": False}}),
        ("text-shadow: 0px -0em 0px #CCC", {"font": {"shadow": False}}),
        ("text-shadow: 0px -0em 0px #999", {"font": {"shadow": False}}),
        ("text-shadow: 0px -0em 0px", {"font": {"shadow": False}}),
        ("text-shadow: 2px -0em 0px #CCC", {"font": {"shadow": True}}),
        ("text-shadow: 0px -2em 0px #CCC", {"font": {"shadow": True}}),
        ("text-shadow: 0px -0em 2px #CCC", {"font": {"shadow": True}}),
        ("text-shadow: 0px -0em 2px", {"font": {"shadow": True}}),
        ("text-shadow: 0px -2em", {"font": {"shadow": True}}),
        # FILL
        # - color, fillType
        (
            "background-color: red",
            {"fill": {"fgColor": "FF0000", "patternType": "solid"}},
        ),
        (
            "background-color: #ff0000",
            {"fill": {"fgColor": "FF0000", "patternType": "solid"}},
        ),
        (
            "background-color: #f0a",
            {"fill": {"fgColor": "FF00AA", "patternType": "solid"}},
        ),
        # BORDER
        # - style
        (
            "border-style: solid",
            {
                "border": {
                    "top": {"style": "medium"},
                    "bottom": {"style": "medium"},
                    "left": {"style": "medium"},
                    "right": {"style": "medium"},
                }
            },
        ),
        (
            "border-style: solid; border-width: thin",
            {
                "border": {
                    "top": {"style": "thin"},
                    "bottom": {"style": "thin"},
                    "left": {"style": "thin"},
                    "right": {"style": "thin"},
                }
            },
        ),
        (
            "border-top-style: solid; border-top-width: thin",
            {"border": {"top": {"style": "thin"}}},
        ),
        (
            "border-top-style: solid; border-top-width: 1pt",
            {"border": {"top": {"style": "thin"}}},
        ),
        ("border-top-style: solid", {"border": {"top": {"style": "medium"}}}),
        (
            "border-top-style: solid; border-top-width: medium",
            {"border": {"top": {"style": "medium"}}},
        ),
        (
            "border-top-style: solid; border-top-width: 2pt",
            {"border": {"top": {"style": "medium"}}},
        ),
        (
            "border-top-style: solid; border-top-width: thick",
            {"border": {"top": {"style": "thick"}}},
        ),
        (
            "border-top-style: solid; border-top-width: 4pt",
            {"border": {"top": {"style": "thick"}}},
        ),
        (
            "border-top-style: dotted",
            {"border": {"top": {"style": "mediumDashDotDot"}}},
        ),
        (
            "border-top-style: dotted; border-top-width: thin",
            {"border": {"top": {"style": "dotted"}}},
        ),
        ("border-top-style: dashed", {"border": {"top": {"style": "mediumDashed"}}}),
        (
            "border-top-style: dashed; border-top-width: thin",
            {"border": {"top": {"style": "dashed"}}},
        ),
        ("border-top-style: double", {"border": {"top": {"style": "double"}}}),
        # - color
        (
            "border-style: solid; border-color: #0000ff",
            {
                "border": {
                    "top": {"style": "medium", "color": "0000FF"},
                    "right": {"style": "medium", "color": "0000FF"},
                    "bottom": {"style": "medium", "color": "0000FF"},
                    "left": {"style": "medium", "color": "0000FF"},
                }
            },
        ),
        (
            "border-top-style: double; border-top-color: blue",
            {"border": {"top": {"style": "double", "color": "0000FF"}}},
        ),
        (
            "border-top-style: solid; border-top-color: #06c",
            {"border": {"top": {"style": "medium", "color": "0066CC"}}},
        ),
        (
            "border-top-color: blue",
            {"border": {"top": {"color": "0000FF", "style": "none"}}},
        ),
        # ALIGNMENT
        # - horizontal
        ("text-align: center", {"alignment": {"horizontal": "center"}}),
        ("text-align: left", {"alignment": {"horizontal": "left"}}),
        ("text-align: right", {"alignment": {"horizontal": "right"}}),
        ("text-align: justify", {"alignment": {"horizontal": "justify"}}),
        # - vertical
        ("vertical-align: top", {"alignment": {"vertical": "top"}}),
        ("vertical-align: text-top", {"alignment": {"vertical": "top"}}),
        ("vertical-align: middle", {"alignment": {"vertical": "center"}}),
        ("vertical-align: bottom", {"alignment": {"vertical": "bottom"}}),
        ("vertical-align: text-bottom", {"alignment": {"vertical": "bottom"}}),
        # - wrap_text
        ("white-space: nowrap", {"alignment": {"wrap_text": False}}),
        ("white-space: pre", {"alignment": {"wrap_text": False}}),
        ("white-space: pre-line", {"alignment": {"wrap_text": False}}),
        ("white-space: normal", {"alignment": {"wrap_text": True}}),
        # NUMBER FORMAT
        ("number-format: 0%", {"number_format": {"format_code": "0%"}}),
        (
            "number-format: 0§[Red](0)§-§@;",
            {"number_format": {"format_code": "0;[red](0);-;@"}},  # GH 46152
        ),
    ],
)
def test_css_to_excel(css, expected):
    convert = CSSToExcelConverter()
    assert expected == convert(css)


def test_css_to_excel_multiple():
    convert = CSSToExcelConverter()
    actual = convert(
        """
        font-weight: bold;
        text-decoration: underline;
        color: red;
        border-width: thin;
        text-align: center;
        vertical-align: top;
        unused: something;
    """
    )
    assert {
        "font": {"bold": True, "underline": "single", "color": "FF0000"},
        "border": {
            "top": {"style": "thin"},
            "right": {"style": "thin"},
            "bottom": {"style": "thin"},
            "left": {"style": "thin"},
        },
        "alignment": {"horizontal": "center", "vertical": "top"},
    } == actual


@pytest.mark.parametrize(
    "css,inherited,expected",
    [
        ("font-weight: bold", "", {"font": {"bold": True}}),
        ("", "font-weight: bold", {"font": {"bold": True}}),
        (
            "font-weight: bold",
            "font-style: italic",
            {"font": {"bold": True, "italic": True}},
        ),
        ("font-style: normal", "font-style: italic", {"font": {"italic": False}}),
        ("font-style: inherit", "", {}),
        (
            "font-style: normal; font-style: inherit",
            "font-style: italic",
            {"font": {"italic": True}},
        ),
    ],
)
def test_css_to_excel_inherited(css, inherited, expected):
    convert = CSSToExcelConverter(inherited)
    assert expected == convert(css)


@pytest.mark.parametrize(
    "input_color,output_color",
    (
        list(CSSToExcelConverter.NAMED_COLORS.items())
        + [("#" + rgb, rgb) for rgb in CSSToExcelConverter.NAMED_COLORS.values()]
        + [("#F0F", "FF00FF"), ("#ABC", "AABBCC")]
    ),
)
def test_css_to_excel_good_colors(input_color, output_color):
    # see gh-18392
    css = (
        f"border-top-color: {input_color}; "
        f"border-right-color: {input_color}; "
        f"border-bottom-color: {input_color}; "
        f"border-left-color: {input_color}; "
        f"background-color: {input_color}; "
        f"color: {input_color}"
    )

    expected = {}

    expected["fill"] = {"patternType": "solid", "fgColor": output_color}

    expected["font"] = {"color": output_color}

    expected["border"] = {
        k: {"color": output_color, "style": "none"}
        for k in ("top", "right", "bottom", "left")
    }

    with tm.assert_produces_warning(None):
        convert = CSSToExcelConverter()
        assert expected == convert(css)


@pytest.mark.parametrize("input_color", [None, "not-a-color"])
def test_css_to_excel_bad_colors(input_color):
    # see gh-18392
    css = (
        f"border-top-color: {input_color}; "
        f"border-right-color: {input_color}; "
        f"border-bottom-color: {input_color}; "
        f"border-left-color: {input_color}; "
        f"background-color: {input_color}; "
        f"color: {input_color}"
    )

    expected = {}

    if input_color is not None:
        expected["fill"] = {"patternType": "solid"}

    with tm.assert_produces_warning(CSSWarning):
        convert = CSSToExcelConverter()
        assert expected == convert(css)


def tests_css_named_colors_valid():
    upper_hexs = set(map(str.upper, string.hexdigits))
    for color in CSSToExcelConverter.NAMED_COLORS.values():
        assert len(color) == 6 and all(c in upper_hexs for c in color)


def test_css_named_colors_from_mpl_present():
    mpl_colors = pytest.importorskip("matplotlib.colors")

    pd_colors = CSSToExcelConverter.NAMED_COLORS
    for name, color in mpl_colors.CSS4_COLORS.items():
        assert name in pd_colors and pd_colors[name] == color[1:]


@pytest.mark.parametrize(
    "styles,expected",
    [
        ([("color", "green"), ("color", "red")], "color: red;"),
        ([("font-weight", "bold"), ("font-weight", "normal")], "font-weight: normal;"),
        ([("text-align", "center"), ("TEXT-ALIGN", "right")], "text-align: right;"),
    ],
)
def test_css_excel_cell_precedence(styles, expected):
    """It applies favors latter declarations over former declarations"""
    # See GH 47371
    converter = CSSToExcelConverter()
    converter._call_cached.cache_clear()
    css_styles = {(0, 0): styles}
    cell = CssExcelCell(
        row=0,
        col=0,
        val="",
        style=None,
        css_styles=css_styles,
        css_row=0,
        css_col=0,
        css_converter=converter,
    )
    converter._call_cached.cache_clear()

    assert cell.style == converter(expected)


@pytest.mark.parametrize(
    "styles,cache_hits,cache_misses",
    [
        ([[("color", "green"), ("color", "red"), ("color", "green")]], 0, 1),
        (
            [
                [("font-weight", "bold")],
                [("font-weight", "normal"), ("font-weight", "bold")],
            ],
            1,
            1,
        ),
        ([[("text-align", "center")], [("TEXT-ALIGN", "center")]], 1, 1),
        (
            [
                [("font-weight", "bold"), ("text-align", "center")],
                [("font-weight", "bold"), ("text-align", "left")],
            ],
            0,
            2,
        ),
        (
            [
                [("font-weight", "bold"), ("text-align", "center")],
                [("font-weight", "bold"), ("text-align", "left")],
                [("font-weight", "bold"), ("text-align", "center")],
            ],
            1,
            2,
        ),
    ],
)
def test_css_excel_cell_cache(styles, cache_hits, cache_misses):
    """It caches unique cell styles"""
    # See GH 47371
    converter = CSSToExcelConverter()
    converter._call_cached.cache_clear()

    css_styles = {(0, i): _style for i, _style in enumerate(styles)}
    for css_row, css_col in css_styles:
        CssExcelCell(
            row=0,
            col=0,
            val="",
            style=None,
            css_styles=css_styles,
            css_row=css_row,
            css_col=css_col,
            css_converter=converter,
        )
    cache_info = converter._call_cached.cache_info()
    converter._call_cached.cache_clear()

    assert cache_info.hits == cache_hits
    assert cache_info.misses == cache_misses

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\matrices\tests\test_sdm.py ===
"""
Tests for the basic functionality of the SDM class.
"""

from itertools import product

from sympy.core.singleton import S
from sympy.external.gmpy import GROUND_TYPES
from sympy.testing.pytest import raises

from sympy.polys.domains import QQ, ZZ, EXRAW
from sympy.polys.matrices.sdm import SDM
from sympy.polys.matrices.ddm import DDM
from sympy.polys.matrices.exceptions import (DMBadInputError, DMDomainError,
                                             DMShapeError)


def test_SDM():
    A = SDM({0:{0:ZZ(1)}}, (2, 2), ZZ)
    assert A.domain == ZZ
    assert A.shape == (2, 2)
    assert dict(A) == {0:{0:ZZ(1)}}

    raises(DMBadInputError, lambda: SDM({5:{1:ZZ(0)}}, (2, 2), ZZ))
    raises(DMBadInputError, lambda: SDM({0:{5:ZZ(0)}}, (2, 2), ZZ))


def test_DDM_str():
    sdm = SDM({0:{0:ZZ(1)}, 1:{1:ZZ(1)}}, (2, 2), ZZ)
    assert str(sdm) == '{0: {0: 1}, 1: {1: 1}}'
    if GROUND_TYPES == 'gmpy': # pragma: no cover
        assert repr(sdm) == 'SDM({0: {0: mpz(1)}, 1: {1: mpz(1)}}, (2, 2), ZZ)'
    else:        # pragma: no cover
        assert repr(sdm) == 'SDM({0: {0: 1}, 1: {1: 1}}, (2, 2), ZZ)'


def test_SDM_new():
    A = SDM({0:{0:ZZ(1)}}, (2, 2), ZZ)
    B = A.new({}, (2, 2), ZZ)
    assert B == SDM({}, (2, 2), ZZ)


def test_SDM_copy():
    A = SDM({0:{0:ZZ(1)}}, (2, 2), ZZ)
    B = A.copy()
    assert A == B
    A[0][0] = ZZ(2)
    assert A != B


def test_SDM_from_list():
    A = SDM.from_list([[ZZ(0), ZZ(1)], [ZZ(1), ZZ(0)]], (2, 2), ZZ)
    assert A == SDM({0:{1:ZZ(1)}, 1:{0:ZZ(1)}}, (2, 2), ZZ)

    raises(DMBadInputError, lambda: SDM.from_list([[ZZ(0)], [ZZ(0), ZZ(1)]], (2, 2), ZZ))
    raises(DMBadInputError, lambda: SDM.from_list([[ZZ(0), ZZ(1)]], (2, 2), ZZ))


def test_SDM_to_list():
    A = SDM({0:{1: ZZ(1)}}, (2, 2), ZZ)
    assert A.to_list() == [[ZZ(0), ZZ(1)], [ZZ(0), ZZ(0)]]

    A = SDM({}, (0, 2), ZZ)
    assert A.to_list() == []

    A = SDM({}, (2, 0), ZZ)
    assert A.to_list() == [[], []]


def test_SDM_to_list_flat():
    A = SDM({0:{1: ZZ(1)}}, (2, 2), ZZ)
    assert A.to_list_flat() == [ZZ(0), ZZ(1), ZZ(0), ZZ(0)]


def test_SDM_to_dok():
    A = SDM({0:{1: ZZ(1)}}, (2, 2), ZZ)
    assert A.to_dok() == {(0, 1): ZZ(1)}


def test_SDM_from_ddm():
    A = DDM([[ZZ(1), ZZ(0)], [ZZ(1), ZZ(0)]], (2, 2), ZZ)
    B = SDM.from_ddm(A)
    assert B.domain == ZZ
    assert B.shape == (2, 2)
    assert dict(B) == {0:{0:ZZ(1)}, 1:{0:ZZ(1)}}


def test_SDM_to_ddm():
    A = SDM({0:{1: ZZ(1)}}, (2, 2), ZZ)
    B = DDM([[ZZ(0), ZZ(1)], [ZZ(0), ZZ(0)]], (2, 2), ZZ)
    assert A.to_ddm() == B


def test_SDM_to_sdm():
    A = SDM({0:{1: ZZ(1)}}, (2, 2), ZZ)
    assert A.to_sdm() == A


def test_SDM_getitem():
    A = SDM({0:{1:ZZ(1)}}, (2, 2), ZZ)
    assert A.getitem(0, 0) == ZZ.zero
    assert A.getitem(0, 1) == ZZ.one
    assert A.getitem(1, 0) == ZZ.zero
    assert A.getitem(-2, -2) == ZZ.zero
    assert A.getitem(-2, -1) == ZZ.one
    assert A.getitem(-1, -2) == ZZ.zero
    raises(IndexError, lambda: A.getitem(2, 0))
    raises(IndexError, lambda: A.getitem(0, 2))


def test_SDM_setitem():
    A = SDM({0:{1:ZZ(1)}}, (2, 2), ZZ)
    A.setitem(0, 0, ZZ(1))
    assert A == SDM({0:{0:ZZ(1), 1:ZZ(1)}}, (2, 2), ZZ)
    A.setitem(1, 0, ZZ(1))
    assert A == SDM({0:{0:ZZ(1), 1:ZZ(1)}, 1:{0:ZZ(1)}}, (2, 2), ZZ)
    A.setitem(1, 0, ZZ(0))
    assert A == SDM({0:{0:ZZ(1), 1:ZZ(1)}}, (2, 2), ZZ)
    # Repeat the above test so that this time the row is empty
    A.setitem(1, 0, ZZ(0))
    assert A == SDM({0:{0:ZZ(1), 1:ZZ(1)}}, (2, 2), ZZ)
    A.setitem(0, 0, ZZ(0))
    assert A == SDM({0:{1:ZZ(1)}}, (2, 2), ZZ)
    # This time the row is there but column is empty
    A.setitem(0, 0, ZZ(0))
    assert A == SDM({0:{1:ZZ(1)}}, (2, 2), ZZ)
    raises(IndexError, lambda: A.setitem(2, 0, ZZ(1)))
    raises(IndexError, lambda: A.setitem(0, 2, ZZ(1)))


def test_SDM_extract_slice():
    A = SDM({0:{0:ZZ(1), 1:ZZ(2)}, 1:{0:ZZ(3), 1:ZZ(4)}}, (2, 2), ZZ)
    B = A.extract_slice(slice(1, 2), slice(1, 2))
    assert B == SDM({0:{0:ZZ(4)}}, (1, 1), ZZ)


def test_SDM_extract():
    A = SDM({0:{0:ZZ(1), 1:ZZ(2)}, 1:{0:ZZ(3), 1:ZZ(4)}}, (2, 2), ZZ)
    B = A.extract([1], [1])
    assert B == SDM({0:{0:ZZ(4)}}, (1, 1), ZZ)
    B = A.extract([1, 0], [1, 0])
    assert B == SDM({0:{0:ZZ(4), 1:ZZ(3)}, 1:{0:ZZ(2), 1:ZZ(1)}}, (2, 2), ZZ)
    B = A.extract([1, 1], [1, 1])
    assert B == SDM({0:{0:ZZ(4), 1:ZZ(4)}, 1:{0:ZZ(4), 1:ZZ(4)}}, (2, 2), ZZ)
    B = A.extract([-1], [-1])
    assert B == SDM({0:{0:ZZ(4)}}, (1, 1), ZZ)

    A = SDM({}, (2, 2), ZZ)
    B = A.extract([0, 1, 0], [0, 0])
    assert B == SDM({}, (3, 2), ZZ)

    A = SDM({0:{0:ZZ(1), 1:ZZ(2)}, 1:{0:ZZ(3), 1:ZZ(4)}}, (2, 2), ZZ)
    assert A.extract([], []) == SDM.zeros((0, 0), ZZ)
    assert A.extract([1], []) == SDM.zeros((1, 0), ZZ)
    assert A.extract([], [1]) == SDM.zeros((0, 1), ZZ)

    raises(IndexError, lambda: A.extract([2], [0]))
    raises(IndexError, lambda: A.extract([0], [2]))
    raises(IndexError, lambda: A.extract([-3], [0]))
    raises(IndexError, lambda: A.extract([0], [-3]))


def test_SDM_zeros():
    A = SDM.zeros((2, 2), ZZ)
    assert A.domain == ZZ
    assert A.shape == (2, 2)
    assert dict(A) == {}

def test_SDM_ones():
    A = SDM.ones((1, 2), QQ)
    assert A.domain == QQ
    assert A.shape == (1, 2)
    assert dict(A) == {0:{0:QQ(1), 1:QQ(1)}}

def test_SDM_eye():
    A = SDM.eye((2, 2), ZZ)
    assert A.domain == ZZ
    assert A.shape == (2, 2)
    assert dict(A) == {0:{0:ZZ(1)}, 1:{1:ZZ(1)}}


def test_SDM_diag():
    A = SDM.diag([ZZ(1), ZZ(2)], ZZ, (2, 3))
    assert A == SDM({0:{0:ZZ(1)}, 1:{1:ZZ(2)}}, (2, 3), ZZ)


def test_SDM_transpose():
    A = SDM({0:{0:ZZ(1), 1:ZZ(2)}, 1:{0:ZZ(3), 1:ZZ(4)}}, (2, 2), ZZ)
    B = SDM({0:{0:ZZ(1), 1:ZZ(3)}, 1:{0:ZZ(2), 1:ZZ(4)}}, (2, 2), ZZ)
    assert A.transpose() == B

    A = SDM({0:{1:ZZ(2)}}, (2, 2), ZZ)
    B = SDM({1:{0:ZZ(2)}}, (2, 2), ZZ)
    assert A.transpose() == B

    A = SDM({0:{1:ZZ(2)}}, (1, 2), ZZ)
    B = SDM({1:{0:ZZ(2)}}, (2, 1), ZZ)
    assert A.transpose() == B


def test_SDM_mul():
    A = SDM({0:{0:ZZ(2)}}, (2, 2), ZZ)
    B = SDM({0:{0:ZZ(4)}}, (2, 2), ZZ)
    assert A*ZZ(2) == B
    assert ZZ(2)*A == B

    raises(TypeError, lambda: A*QQ(1, 2))
    raises(TypeError, lambda: QQ(1, 2)*A)


def test_SDM_mul_elementwise():
    A = SDM({0:{0:ZZ(2), 1:ZZ(2)}}, (2, 2), ZZ)
    B = SDM({0:{0:ZZ(4)}, 1:{0:ZZ(3)}}, (2, 2), ZZ)
    C = SDM({0:{0:ZZ(8)}}, (2, 2), ZZ)
    assert A.mul_elementwise(B) == C
    assert B.mul_elementwise(A) == C

    Aq = A.convert_to(QQ)
    A1 = SDM({0:{0:ZZ(1)}}, (1, 1), ZZ)

    raises(DMDomainError, lambda: Aq.mul_elementwise(B))
    raises(DMShapeError, lambda: A1.mul_elementwise(B))


def test_SDM_matmul():
    A = SDM({0:{0:ZZ(2)}}, (2, 2), ZZ)
    B = SDM({0:{0:ZZ(4)}}, (2, 2), ZZ)
    assert A.matmul(A) == A*A == B

    C = SDM({0:{0:ZZ(2)}}, (2, 2), QQ)
    raises(DMDomainError, lambda: A.matmul(C))

    A = SDM({0:{0:ZZ(1), 1:ZZ(2)}, 1:{0:ZZ(3), 1:ZZ(4)}}, (2, 2), ZZ)
    B = SDM({0:{0:ZZ(7), 1:ZZ(10)}, 1:{0:ZZ(15), 1:ZZ(22)}}, (2, 2), ZZ)
    assert A.matmul(A) == A*A == B

    A22 = SDM({0:{0:ZZ(4)}}, (2, 2), ZZ)
    A32 = SDM({0:{0:ZZ(2)}}, (3, 2), ZZ)
    A23 = SDM({0:{0:ZZ(4)}}, (2, 3), ZZ)
    A33 = SDM({0:{0:ZZ(8)}}, (3, 3), ZZ)
    A22 = SDM({0:{0:ZZ(8)}}, (2, 2), ZZ)
    assert A32.matmul(A23) == A33
    assert A23.matmul(A32) == A22
    # XXX: @ not supported by SDM...
    #assert A32.matmul(A23) == A32 @ A23 == A33
    #assert A23.matmul(A32) == A23 @ A32 == A22
    #raises(DMShapeError, lambda: A23 @ A22)
    raises(DMShapeError, lambda: A23.matmul(A22))

    A = SDM({0: {0: ZZ(-1), 1: ZZ(1)}}, (1, 2), ZZ)
    B = SDM({0: {0: ZZ(-1)}, 1: {0: ZZ(-1)}}, (2, 1), ZZ)
    assert A.matmul(B) == A*B == SDM({}, (1, 1), ZZ)


def test_matmul_exraw():

    def dm(d):
        result = {}
        for i, row in d.items():
            row = {j:val for j, val in row.items() if val}
            if row:
                result[i] = row
        return SDM(result, (2, 2), EXRAW)

    values = [S.NegativeInfinity, S.NegativeOne, S.Zero, S.One, S.Infinity]
    for a, b, c, d in product(*[values]*4):
        Ad = dm({0: {0:a, 1:b}, 1: {0:c, 1:d}})
        Ad2 = dm({0: {0:a*a + b*c, 1:a*b + b*d}, 1:{0:c*a + d*c, 1: c*b + d*d}})
        assert Ad * Ad == Ad2


def test_SDM_add():
    A = SDM({0:{1:ZZ(1)}, 1:{0:ZZ(2), 1:ZZ(3)}}, (2, 2), ZZ)
    B = SDM({0:{0:ZZ(1)}, 1:{0:ZZ(-2), 1:ZZ(3)}}, (2, 2), ZZ)
    C = SDM({0:{0:ZZ(1), 1:ZZ(1)}, 1:{1:ZZ(6)}}, (2, 2), ZZ)
    assert A.add(B) == B.add(A) == A + B == B + A == C

    A = SDM({0:{1:ZZ(1)}}, (2, 2), ZZ)
    B = SDM({0:{0:ZZ(1)}, 1:{0:ZZ(-2), 1:ZZ(3)}}, (2, 2), ZZ)
    C = SDM({0:{0:ZZ(1), 1:ZZ(1)}, 1:{0:ZZ(-2), 1:ZZ(3)}}, (2, 2), ZZ)
    assert A.add(B) == B.add(A) == A + B == B + A == C

    raises(TypeError, lambda: A + [])


def test_SDM_sub():
    A = SDM({0:{1:ZZ(1)}, 1:{0:ZZ(2), 1:ZZ(3)}}, (2, 2), ZZ)
    B = SDM({0:{0:ZZ(1)}, 1:{0:ZZ(-2), 1:ZZ(3)}}, (2, 2), ZZ)
    C = SDM({0:{0:ZZ(-1), 1:ZZ(1)}, 1:{0:ZZ(4)}}, (2, 2), ZZ)
    assert A.sub(B) == A - B == C

    raises(TypeError, lambda: A - [])


def test_SDM_neg():
    A = SDM({0:{1:ZZ(1)}, 1:{0:ZZ(2), 1:ZZ(3)}}, (2, 2), ZZ)
    B = SDM({0:{1:ZZ(-1)}, 1:{0:ZZ(-2), 1:ZZ(-3)}}, (2, 2), ZZ)
    assert A.neg() == -A == B


def test_SDM_convert_to():
    A = SDM({0:{1:ZZ(1)}, 1:{0:ZZ(2), 1:ZZ(3)}}, (2, 2), ZZ)
    B = SDM({0:{1:QQ(1)}, 1:{0:QQ(2), 1:QQ(3)}}, (2, 2), QQ)
    C = A.convert_to(QQ)
    assert C == B
    assert C.domain == QQ

    D = A.convert_to(ZZ)
    assert D == A
    assert D.domain == ZZ


def test_SDM_hstack():
    A = SDM({0:{1:ZZ(1)}}, (2, 2), ZZ)
    B = SDM({1:{1:ZZ(1)}}, (2, 2), ZZ)
    AA = SDM({0:{1:ZZ(1), 3:ZZ(1)}}, (2, 4), ZZ)
    AB = SDM({0:{1:ZZ(1)}, 1:{3:ZZ(1)}}, (2, 4), ZZ)
    assert SDM.hstack(A) == A
    assert SDM.hstack(A, A) == AA
    assert SDM.hstack(A, B) == AB


def test_SDM_vstack():
    A = SDM({0:{1:ZZ(1)}}, (2, 2), ZZ)
    B = SDM({1:{1:ZZ(1)}}, (2, 2), ZZ)
    AA = SDM({0:{1:ZZ(1)}, 2:{1:ZZ(1)}}, (4, 2), ZZ)
    AB = SDM({0:{1:ZZ(1)}, 3:{1:ZZ(1)}}, (4, 2), ZZ)
    assert SDM.vstack(A) == A
    assert SDM.vstack(A, A) == AA
    assert SDM.vstack(A, B) == AB


def test_SDM_applyfunc():
    A = SDM({0:{1:ZZ(1)}}, (2, 2), ZZ)
    B = SDM({0:{1:ZZ(2)}}, (2, 2), ZZ)
    assert A.applyfunc(lambda x: 2*x, ZZ) == B


def test_SDM_inv():
    A = SDM({0:{0:QQ(1), 1:QQ(2)}, 1:{0:QQ(3), 1:QQ(4)}}, (2, 2), QQ)
    B = SDM({0:{0:QQ(-2), 1:QQ(1)}, 1:{0:QQ(3, 2), 1:QQ(-1, 2)}}, (2, 2), QQ)
    assert A.inv() == B


def test_SDM_det():
    A = SDM({0:{0:QQ(1), 1:QQ(2)}, 1:{0:QQ(3), 1:QQ(4)}}, (2, 2), QQ)
    assert A.det() == QQ(-2)


def test_SDM_lu():
    A = SDM({0:{0:QQ(1), 1:QQ(2)}, 1:{0:QQ(3), 1:QQ(4)}}, (2, 2), QQ)
    L = SDM({0:{0:QQ(1)}, 1:{0:QQ(3), 1:QQ(1)}}, (2, 2), QQ)
    #U = SDM({0:{0:QQ(1), 1:QQ(2)}, 1:{0:QQ(3), 1:QQ(-2)}}, (2, 2), QQ)
    #swaps = []
    # This doesn't quite work. U has some nonzero elements in the lower part.
    #assert A.lu() == (L, U, swaps)
    assert A.lu()[0] == L


def test_SDM_lu_solve():
    A = SDM({0:{0:QQ(1), 1:QQ(2)}, 1:{0:QQ(3), 1:QQ(4)}}, (2, 2), QQ)
    b = SDM({0:{0:QQ(1)}, 1:{0:QQ(2)}}, (2, 1), QQ)
    x = SDM({1:{0:QQ(1, 2)}}, (2, 1), QQ)
    assert A.matmul(x) == b
    assert A.lu_solve(b) == x


def test_SDM_charpoly():
    A = SDM({0:{0:ZZ(1), 1:ZZ(2)}, 1:{0:ZZ(3), 1:ZZ(4)}}, (2, 2), ZZ)
    assert A.charpoly() == [ZZ(1), ZZ(-5), ZZ(-2)]


def test_SDM_nullspace():
    # More tests are in test_nullspace.py
    A = SDM({0:{0:QQ(1), 1:QQ(1)}}, (2, 2), QQ)
    assert A.nullspace()[0] == SDM({0:{0:QQ(-1), 1:QQ(1)}}, (1, 2), QQ)


def test_SDM_rref():
    # More tests are in test_rref.py

    A = SDM({0:{0:QQ(1), 1:QQ(2)},
             1:{0:QQ(3), 1:QQ(4)}}, (2, 2), QQ)
    A_rref = SDM({0:{0:QQ(1)}, 1:{1:QQ(1)}}, (2, 2), QQ)
    assert A.rref() == (A_rref, [0, 1])

    A = SDM({0: {0: QQ(1), 1: QQ(2), 2: QQ(2)},
             1: {0: QQ(3),           2: QQ(4)}}, (2, 3), ZZ)
    A_rref = SDM({0: {0: QQ(1,1), 2: QQ(4,3)},
                  1: {1: QQ(1,1), 2: QQ(1,3)}}, (2, 3), QQ)
    assert A.rref() == (A_rref, [0, 1])


def test_SDM_particular():
    A = SDM({0:{0:QQ(1)}}, (2, 2), QQ)
    Apart = SDM.zeros((1, 2), QQ)
    assert A.particular() == Apart


def test_SDM_is_zero_matrix():
    A = SDM({0: {0: QQ(1)}}, (2, 2), QQ)
    Azero = SDM.zeros((1, 2), QQ)
    assert A.is_zero_matrix() is False
    assert Azero.is_zero_matrix() is True


def test_SDM_is_upper():
    A = SDM({0: {0: QQ(1), 1: QQ(2), 2: QQ(3), 3: QQ(4)},
                       1: {1: QQ(5), 2: QQ(6), 3: QQ(7)},
                                 2: {2: QQ(8), 3: QQ(9)}}, (3, 4), QQ)
    B = SDM({0: {0: QQ(1), 1: QQ(2), 2: QQ(3), 3: QQ(4)},
                       1: {1: QQ(5), 2: QQ(6), 3: QQ(7)},
                       2: {1: QQ(7), 2: QQ(8), 3: QQ(9)}}, (3, 4), QQ)
    assert A.is_upper() is True
    assert B.is_upper() is False


def test_SDM_is_lower():
    A = SDM({0: {0: QQ(1), 1: QQ(2), 2: QQ(3), 3: QQ(4)},
                       1: {1: QQ(5), 2: QQ(6), 3: QQ(7)},
                                 2: {2: QQ(8), 3: QQ(9)}}, (3, 4), QQ
            ).transpose()
    B = SDM({0: {0: QQ(1), 1: QQ(2), 2: QQ(3), 3: QQ(4)},
                       1: {1: QQ(5), 2: QQ(6), 3: QQ(7)},
                       2: {1: QQ(7), 2: QQ(8), 3: QQ(9)}}, (3, 4), QQ
            ).transpose()
    assert A.is_lower() is True
    assert B.is_lower() is False

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\strings\test_cat.py ===
import re

import numpy as np
import pytest

import pandas.util._test_decorators as td

from pandas import (
    DataFrame,
    Index,
    MultiIndex,
    Series,
    _testing as tm,
    concat,
    option_context,
)


@pytest.mark.parametrize("other", [None, Series, Index])
def test_str_cat_name(index_or_series, other):
    # GH 21053
    box = index_or_series
    values = ["a", "b"]
    if other:
        other = other(values)
    else:
        other = values
    result = box(values, name="name").str.cat(other, sep=",")
    assert result.name == "name"


@pytest.mark.parametrize(
    "infer_string", [False, pytest.param(True, marks=td.skip_if_no("pyarrow"))]
)
def test_str_cat(index_or_series, infer_string):
    with option_context("future.infer_string", infer_string):
        box = index_or_series
        # test_cat above tests "str_cat" from ndarray;
        # here testing "str.cat" from Series/Index to ndarray/list
        s = box(["a", "a", "b", "b", "c", np.nan])

        # single array
        result = s.str.cat()
        expected = "aabbc"
        assert result == expected

        result = s.str.cat(na_rep="-")
        expected = "aabbc-"
        assert result == expected

        result = s.str.cat(sep="_", na_rep="NA")
        expected = "a_a_b_b_c_NA"
        assert result == expected

        t = np.array(["a", np.nan, "b", "d", "foo", np.nan], dtype=object)
        expected = box(["aa", "a-", "bb", "bd", "cfoo", "--"])

        # Series/Index with array
        result = s.str.cat(t, na_rep="-")
        tm.assert_equal(result, expected)

        # Series/Index with list
        result = s.str.cat(list(t), na_rep="-")
        tm.assert_equal(result, expected)

        # errors for incorrect lengths
        rgx = r"If `others` contains arrays or lists \(or other list-likes.*"
        z = Series(["1", "2", "3"])

        with pytest.raises(ValueError, match=rgx):
            s.str.cat(z.values)

        with pytest.raises(ValueError, match=rgx):
            s.str.cat(list(z))


def test_str_cat_raises_intuitive_error(index_or_series):
    # GH 11334
    box = index_or_series
    s = box(["a", "b", "c", "d"])
    message = "Did you mean to supply a `sep` keyword?"
    with pytest.raises(ValueError, match=message):
        s.str.cat("|")
    with pytest.raises(ValueError, match=message):
        s.str.cat("    ")


@pytest.mark.parametrize(
    "infer_string", [False, pytest.param(True, marks=td.skip_if_no("pyarrow"))]
)
@pytest.mark.parametrize("sep", ["", None])
@pytest.mark.parametrize("dtype_target", ["object", "category"])
@pytest.mark.parametrize("dtype_caller", ["object", "category"])
def test_str_cat_categorical(
    index_or_series, dtype_caller, dtype_target, sep, infer_string
):
    box = index_or_series

    with option_context("future.infer_string", infer_string):
        s = Index(["a", "a", "b", "a"], dtype=dtype_caller)
        s = s if box == Index else Series(s, index=s, dtype=s.dtype)
        t = Index(["b", "a", "b", "c"], dtype=dtype_target)

        expected = Index(
            ["ab", "aa", "bb", "ac"], dtype=object if dtype_caller == "object" else None
        )
        expected = (
            expected
            if box == Index
            else Series(
                expected, index=Index(s, dtype=dtype_caller), dtype=expected.dtype
            )
        )

        # Series/Index with unaligned Index -> t.values
        result = s.str.cat(t.values, sep=sep)
        tm.assert_equal(result, expected)

        # Series/Index with Series having matching Index
        t = Series(t.values, index=Index(s, dtype=dtype_caller))
        result = s.str.cat(t, sep=sep)
        tm.assert_equal(result, expected)

        # Series/Index with Series.values
        result = s.str.cat(t.values, sep=sep)
        tm.assert_equal(result, expected)

        # Series/Index with Series having different Index
        t = Series(t.values, index=t.values)
        expected = Index(
            ["aa", "aa", "bb", "bb", "aa"],
            dtype=object if dtype_caller == "object" else None,
        )
        dtype = object if dtype_caller == "object" else s.dtype.categories.dtype
        expected = (
            expected
            if box == Index
            else Series(
                expected,
                index=Index(expected.str[:1], dtype=dtype),
                dtype=expected.dtype,
            )
        )

        result = s.str.cat(t, sep=sep)
        tm.assert_equal(result, expected)


@pytest.mark.parametrize(
    "data",
    [[1, 2, 3], [0.1, 0.2, 0.3], [1, 2, "b"]],
    ids=["integers", "floats", "mixed"],
)
# without dtype=object, np.array would cast [1, 2, 'b'] to ['1', '2', 'b']
@pytest.mark.parametrize(
    "box",
    [Series, Index, list, lambda x: np.array(x, dtype=object)],
    ids=["Series", "Index", "list", "np.array"],
)
def test_str_cat_wrong_dtype_raises(box, data):
    # GH 22722
    s = Series(["a", "b", "c"])
    t = box(data)

    msg = "Concatenation requires list-likes containing only strings.*"
    with pytest.raises(TypeError, match=msg):
        # need to use outer and na_rep, as otherwise Index would not raise
        s.str.cat(t, join="outer", na_rep="-")


def test_str_cat_mixed_inputs(index_or_series):
    box = index_or_series
    s = Index(["a", "b", "c", "d"])
    s = s if box == Index else Series(s, index=s)

    t = Series(["A", "B", "C", "D"], index=s.values)
    d = concat([t, Series(s, index=s)], axis=1)

    expected = Index(["aAa", "bBb", "cCc", "dDd"])
    expected = expected if box == Index else Series(expected.values, index=s.values)

    # Series/Index with DataFrame
    result = s.str.cat(d)
    tm.assert_equal(result, expected)

    # Series/Index with two-dimensional ndarray
    result = s.str.cat(d.values)
    tm.assert_equal(result, expected)

    # Series/Index with list of Series
    result = s.str.cat([t, s])
    tm.assert_equal(result, expected)

    # Series/Index with mixed list of Series/array
    result = s.str.cat([t, s.values])
    tm.assert_equal(result, expected)

    # Series/Index with list of Series; different indexes
    t.index = ["b", "c", "d", "a"]
    expected = box(["aDa", "bAb", "cBc", "dCd"])
    expected = expected if box == Index else Series(expected.values, index=s.values)
    result = s.str.cat([t, s])
    tm.assert_equal(result, expected)

    # Series/Index with mixed list; different index
    result = s.str.cat([t, s.values])
    tm.assert_equal(result, expected)

    # Series/Index with DataFrame; different indexes
    d.index = ["b", "c", "d", "a"]
    expected = box(["aDd", "bAa", "cBb", "dCc"])
    expected = expected if box == Index else Series(expected.values, index=s.values)
    result = s.str.cat(d)
    tm.assert_equal(result, expected)

    # errors for incorrect lengths
    rgx = r"If `others` contains arrays or lists \(or other list-likes.*"
    z = Series(["1", "2", "3"])
    e = concat([z, z], axis=1)

    # two-dimensional ndarray
    with pytest.raises(ValueError, match=rgx):
        s.str.cat(e.values)

    # list of list-likes
    with pytest.raises(ValueError, match=rgx):
        s.str.cat([z.values, s.values])

    # mixed list of Series/list-like
    with pytest.raises(ValueError, match=rgx):
        s.str.cat([z.values, s])

    # errors for incorrect arguments in list-like
    rgx = "others must be Series, Index, DataFrame,.*"
    # make sure None/NaN do not crash checks in _get_series_list
    u = Series(["a", np.nan, "c", None])

    # mix of string and Series
    with pytest.raises(TypeError, match=rgx):
        s.str.cat([u, "u"])

    # DataFrame in list
    with pytest.raises(TypeError, match=rgx):
        s.str.cat([u, d])

    # 2-dim ndarray in list
    with pytest.raises(TypeError, match=rgx):
        s.str.cat([u, d.values])

    # nested lists
    with pytest.raises(TypeError, match=rgx):
        s.str.cat([u, [u, d]])

    # forbidden input type: set
    # GH 23009
    with pytest.raises(TypeError, match=rgx):
        s.str.cat(set(u))

    # forbidden input type: set in list
    # GH 23009
    with pytest.raises(TypeError, match=rgx):
        s.str.cat([u, set(u)])

    # other forbidden input type, e.g. int
    with pytest.raises(TypeError, match=rgx):
        s.str.cat(1)

    # nested list-likes
    with pytest.raises(TypeError, match=rgx):
        s.str.cat(iter([t.values, list(s)]))


@pytest.mark.parametrize("join", ["left", "outer", "inner", "right"])
def test_str_cat_align_indexed(index_or_series, join):
    # https://github.com/pandas-dev/pandas/issues/18657
    box = index_or_series

    s = Series(["a", "b", "c", "d"], index=["a", "b", "c", "d"])
    t = Series(["D", "A", "E", "B"], index=["d", "a", "e", "b"])
    sa, ta = s.align(t, join=join)
    # result after manual alignment of inputs
    expected = sa.str.cat(ta, na_rep="-")

    if box == Index:
        s = Index(s)
        sa = Index(sa)
        expected = Index(expected)

    result = s.str.cat(t, join=join, na_rep="-")
    tm.assert_equal(result, expected)


@pytest.mark.parametrize("join", ["left", "outer", "inner", "right"])
def test_str_cat_align_mixed_inputs(join):
    s = Series(["a", "b", "c", "d"])
    t = Series(["d", "a", "e", "b"], index=[3, 0, 4, 1])
    d = concat([t, t], axis=1)

    expected_outer = Series(["aaa", "bbb", "c--", "ddd", "-ee"])
    expected = expected_outer.loc[s.index.join(t.index, how=join)]

    # list of Series
    result = s.str.cat([t, t], join=join, na_rep="-")
    tm.assert_series_equal(result, expected)

    # DataFrame
    result = s.str.cat(d, join=join, na_rep="-")
    tm.assert_series_equal(result, expected)

    # mixed list of indexed/unindexed
    u = np.array(["A", "B", "C", "D"])
    expected_outer = Series(["aaA", "bbB", "c-C", "ddD", "-e-"])
    # joint index of rhs [t, u]; u will be forced have index of s
    rhs_idx = (
        t.index.intersection(s.index)
        if join == "inner"
        else t.index.union(s.index)
        if join == "outer"
        else t.index.append(s.index.difference(t.index))
    )

    expected = expected_outer.loc[s.index.join(rhs_idx, how=join)]
    result = s.str.cat([t, u], join=join, na_rep="-")
    tm.assert_series_equal(result, expected)

    with pytest.raises(TypeError, match="others must be Series,.*"):
        # nested lists are forbidden
        s.str.cat([t, list(u)], join=join)

    # errors for incorrect lengths
    rgx = r"If `others` contains arrays or lists \(or other list-likes.*"
    z = Series(["1", "2", "3"]).values

    # unindexed object of wrong length
    with pytest.raises(ValueError, match=rgx):
        s.str.cat(z, join=join)

    # unindexed object of wrong length in list
    with pytest.raises(ValueError, match=rgx):
        s.str.cat([t, z], join=join)


def test_str_cat_all_na(index_or_series, index_or_series2):
    # GH 24044
    box = index_or_series
    other = index_or_series2

    # check that all NaNs in caller / target work
    s = Index(["a", "b", "c", "d"])
    s = s if box == Index else Series(s, index=s)
    t = other([np.nan] * 4, dtype=object)
    # add index of s for alignment
    t = t if other == Index else Series(t, index=s)

    # all-NA target
    if box == Series:
        expected = Series([np.nan] * 4, index=s.index, dtype=s.dtype)
    else:  # box == Index
        # TODO: Strimg option, this should return string dtype
        expected = Index([np.nan] * 4, dtype=object)
    result = s.str.cat(t, join="left")
    tm.assert_equal(result, expected)

    # all-NA caller (only for Series)
    if other == Series:
        expected = Series([np.nan] * 4, dtype=object, index=t.index)
        result = t.str.cat(s, join="left")
        tm.assert_series_equal(result, expected)


def test_str_cat_special_cases():
    s = Series(["a", "b", "c", "d"])
    t = Series(["d", "a", "e", "b"], index=[3, 0, 4, 1])

    # iterator of elements with different types
    expected = Series(["aaa", "bbb", "c-c", "ddd", "-e-"])
    result = s.str.cat(iter([t, s.values]), join="outer", na_rep="-")
    tm.assert_series_equal(result, expected)

    # right-align with different indexes in others
    expected = Series(["aa-", "d-d"], index=[0, 3])
    result = s.str.cat([t.loc[[0]], t.loc[[3]]], join="right", na_rep="-")
    tm.assert_series_equal(result, expected)


def test_cat_on_filtered_index():
    df = DataFrame(
        index=MultiIndex.from_product(
            [[2011, 2012], [1, 2, 3]], names=["year", "month"]
        )
    )

    df = df.reset_index()
    df = df[df.month > 1]

    str_year = df.year.astype("str")
    str_month = df.month.astype("str")
    str_both = str_year.str.cat(str_month, sep=" ")

    assert str_both.loc[1] == "2011 2"

    str_multiple = str_year.str.cat([str_month, str_month], sep=" ")

    assert str_multiple.loc[1] == "2011 2 2"


@pytest.mark.parametrize("klass", [tuple, list, np.array, Series, Index])
def test_cat_different_classes(klass):
    # https://github.com/pandas-dev/pandas/issues/33425
    s = Series(["a", "b", "c"])
    result = s.str.cat(klass(["x", "y", "z"]))
    expected = Series(["ax", "by", "cz"])
    tm.assert_series_equal(result, expected)


def test_cat_on_series_dot_str():
    # GH 28277
    ps = Series(["AbC", "de", "FGHI", "j", "kLLLm"])

    message = re.escape(
        "others must be Series, Index, DataFrame, np.ndarray "
        "or list-like (either containing only strings or "
        "containing only objects of type Series/Index/"
        "np.ndarray[1-dim])"
    )
    with pytest.raises(TypeError, match=message):
        ps.str.cat(others=ps.str)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\hep\tests\test_gamma_matrices.py ===
from sympy.matrices.dense import eye, Matrix
from sympy.tensor.tensor import tensor_indices, TensorHead, tensor_heads, \
    TensExpr, canon_bp
from sympy.physics.hep.gamma_matrices import GammaMatrix as G, LorentzIndex, \
    kahane_simplify, gamma_trace, _simplify_single_line, simplify_gamma_expression
from sympy import Symbol


def _is_tensor_eq(arg1, arg2):
    arg1 = canon_bp(arg1)
    arg2 = canon_bp(arg2)
    if isinstance(arg1, TensExpr):
        return arg1.equals(arg2)
    elif isinstance(arg2, TensExpr):
        return arg2.equals(arg1)
    return arg1 == arg2

def execute_gamma_simplify_tests_for_function(tfunc, D):
    """
    Perform tests to check if sfunc is able to simplify gamma matrix expressions.

    Parameters
    ==========

    `sfunc`     a function to simplify a `TIDS`, shall return the simplified `TIDS`.
    `D`         the number of dimension (in most cases `D=4`).

    """

    mu, nu, rho, sigma = tensor_indices("mu, nu, rho, sigma", LorentzIndex)
    a1, a2, a3, a4, a5, a6 = tensor_indices("a1:7", LorentzIndex)
    mu11, mu12, mu21, mu31, mu32, mu41, mu51, mu52 = tensor_indices("mu11, mu12, mu21, mu31, mu32, mu41, mu51, mu52", LorentzIndex)
    mu61, mu71, mu72 = tensor_indices("mu61, mu71, mu72", LorentzIndex)
    m0, m1, m2, m3, m4, m5, m6 = tensor_indices("m0:7", LorentzIndex)

    def g(xx, yy):
        return (G(xx)*G(yy) + G(yy)*G(xx))/2

    # Some examples taken from Kahane's paper, 4 dim only:
    if D == 4:
        t = (G(a1)*G(mu11)*G(a2)*G(mu21)*G(-a1)*G(mu31)*G(-a2))
        assert _is_tensor_eq(tfunc(t), -4*G(mu11)*G(mu31)*G(mu21) - 4*G(mu31)*G(mu11)*G(mu21))

        t = (G(a1)*G(mu11)*G(mu12)*\
                              G(a2)*G(mu21)*\
                              G(a3)*G(mu31)*G(mu32)*\
                              G(a4)*G(mu41)*\
                              G(-a2)*G(mu51)*G(mu52)*\
                              G(-a1)*G(mu61)*\
                              G(-a3)*G(mu71)*G(mu72)*\
                              G(-a4))
        assert _is_tensor_eq(tfunc(t), \
            16*G(mu31)*G(mu32)*G(mu72)*G(mu71)*G(mu11)*G(mu52)*G(mu51)*G(mu12)*G(mu61)*G(mu21)*G(mu41) + 16*G(mu31)*G(mu32)*G(mu72)*G(mu71)*G(mu12)*G(mu51)*G(mu52)*G(mu11)*G(mu61)*G(mu21)*G(mu41) + 16*G(mu71)*G(mu72)*G(mu32)*G(mu31)*G(mu11)*G(mu52)*G(mu51)*G(mu12)*G(mu61)*G(mu21)*G(mu41) + 16*G(mu71)*G(mu72)*G(mu32)*G(mu31)*G(mu12)*G(mu51)*G(mu52)*G(mu11)*G(mu61)*G(mu21)*G(mu41))

    # Fully Lorentz-contracted expressions, these return scalars:

    def add_delta(ne):
        return ne * eye(4)  # DiracSpinorIndex.delta(DiracSpinorIndex.auto_left, -DiracSpinorIndex.auto_right)

    t = (G(mu)*G(-mu))
    ts = add_delta(D)
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(mu)*G(nu)*G(-mu)*G(-nu))
    ts = add_delta(2*D - D**2)  # -8
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(mu)*G(nu)*G(-nu)*G(-mu))
    ts = add_delta(D**2)  # 16
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(mu)*G(nu)*G(-rho)*G(-nu)*G(-mu)*G(rho))
    ts = add_delta(4*D - 4*D**2 + D**3)  # 16
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(mu)*G(nu)*G(rho)*G(-rho)*G(-nu)*G(-mu))
    ts = add_delta(D**3)  # 64
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(a1)*G(a2)*G(a3)*G(a4)*G(-a3)*G(-a1)*G(-a2)*G(-a4))
    ts = add_delta(-8*D + 16*D**2 - 8*D**3 + D**4)  # -32
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(-mu)*G(-nu)*G(-rho)*G(-sigma)*G(nu)*G(mu)*G(sigma)*G(rho))
    ts = add_delta(-16*D + 24*D**2 - 8*D**3 + D**4)  # 64
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(-mu)*G(nu)*G(-rho)*G(sigma)*G(rho)*G(-nu)*G(mu)*G(-sigma))
    ts = add_delta(8*D - 12*D**2 + 6*D**3 - D**4)  # -32
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(a1)*G(a2)*G(a3)*G(a4)*G(a5)*G(-a3)*G(-a2)*G(-a1)*G(-a5)*G(-a4))
    ts = add_delta(64*D - 112*D**2 + 60*D**3 - 12*D**4 + D**5)  # 256
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(a1)*G(a2)*G(a3)*G(a4)*G(a5)*G(-a3)*G(-a1)*G(-a2)*G(-a4)*G(-a5))
    ts = add_delta(64*D - 120*D**2 + 72*D**3 - 16*D**4 + D**5)  # -128
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(a1)*G(a2)*G(a3)*G(a4)*G(a5)*G(a6)*G(-a3)*G(-a2)*G(-a1)*G(-a6)*G(-a5)*G(-a4))
    ts = add_delta(416*D - 816*D**2 + 528*D**3 - 144*D**4 + 18*D**5 - D**6)  # -128
    assert _is_tensor_eq(tfunc(t), ts)

    t = (G(a1)*G(a2)*G(a3)*G(a4)*G(a5)*G(a6)*G(-a2)*G(-a3)*G(-a1)*G(-a6)*G(-a4)*G(-a5))
    ts = add_delta(416*D - 848*D**2 + 584*D**3 - 172*D**4 + 22*D**5 - D**6)  # -128
    assert _is_tensor_eq(tfunc(t), ts)

    # Expressions with free indices:

    t = (G(mu)*G(nu)*G(rho)*G(sigma)*G(-mu))
    assert _is_tensor_eq(tfunc(t), (-2*G(sigma)*G(rho)*G(nu) + (4-D)*G(nu)*G(rho)*G(sigma)))

    t = (G(mu)*G(nu)*G(-mu))
    assert _is_tensor_eq(tfunc(t), (2-D)*G(nu))

    t = (G(mu)*G(nu)*G(rho)*G(-mu))
    assert _is_tensor_eq(tfunc(t), 2*G(nu)*G(rho) + 2*G(rho)*G(nu) - (4-D)*G(nu)*G(rho))

    t = 2*G(m2)*G(m0)*G(m1)*G(-m0)*G(-m1)
    st = tfunc(t)
    assert _is_tensor_eq(st, (D*(-2*D + 4))*G(m2))

    t = G(m2)*G(m0)*G(m1)*G(-m0)*G(-m2)
    st = tfunc(t)
    assert _is_tensor_eq(st, ((-D + 2)**2)*G(m1))

    t = G(m0)*G(m1)*G(m2)*G(m3)*G(-m1)
    st = tfunc(t)
    assert _is_tensor_eq(st, (D - 4)*G(m0)*G(m2)*G(m3) + 4*G(m0)*g(m2, m3))

    t = G(m0)*G(m1)*G(m2)*G(m3)*G(-m1)*G(-m0)
    st = tfunc(t)
    assert _is_tensor_eq(st, ((D - 4)**2)*G(m2)*G(m3) + (8*D - 16)*g(m2, m3))

    t = G(m2)*G(m0)*G(m1)*G(-m2)*G(-m0)
    st = tfunc(t)
    assert _is_tensor_eq(st, ((-D + 2)*(D - 4) + 4)*G(m1))

    t = G(m3)*G(m1)*G(m0)*G(m2)*G(-m3)*G(-m0)*G(-m2)
    st = tfunc(t)
    assert _is_tensor_eq(st, (-4*D + (-D + 2)**2*(D - 4) + 8)*G(m1))

    t = 2*G(m0)*G(m1)*G(m2)*G(m3)*G(-m0)
    st = tfunc(t)
    assert _is_tensor_eq(st, ((-2*D + 8)*G(m1)*G(m2)*G(m3) - 4*G(m3)*G(m2)*G(m1)))

    t = G(m5)*G(m0)*G(m1)*G(m4)*G(m2)*G(-m4)*G(m3)*G(-m0)
    st = tfunc(t)
    assert _is_tensor_eq(st, (((-D + 2)*(-D + 4))*G(m5)*G(m1)*G(m2)*G(m3) + (2*D - 4)*G(m5)*G(m3)*G(m2)*G(m1)))

    t = -G(m0)*G(m1)*G(m2)*G(m3)*G(-m0)*G(m4)
    st = tfunc(t)
    assert _is_tensor_eq(st, ((D - 4)*G(m1)*G(m2)*G(m3)*G(m4) + 2*G(m3)*G(m2)*G(m1)*G(m4)))

    t = G(-m5)*G(m0)*G(m1)*G(m2)*G(m3)*G(m4)*G(-m0)*G(m5)
    st = tfunc(t)

    result1 = ((-D + 4)**2 + 4)*G(m1)*G(m2)*G(m3)*G(m4) +\
        (4*D - 16)*G(m3)*G(m2)*G(m1)*G(m4) + (4*D - 16)*G(m4)*G(m1)*G(m2)*G(m3)\
        + 4*G(m2)*G(m1)*G(m4)*G(m3) + 4*G(m3)*G(m4)*G(m1)*G(m2) +\
        4*G(m4)*G(m3)*G(m2)*G(m1)

    # Kahane's algorithm yields this result, which is equivalent to `result1`
    # in four dimensions, but is not automatically recognized as equal:
    result2 = 8*G(m1)*G(m2)*G(m3)*G(m4) + 8*G(m4)*G(m3)*G(m2)*G(m1)

    if D == 4:
        assert _is_tensor_eq(st, (result1)) or _is_tensor_eq(st, (result2))
    else:
        assert _is_tensor_eq(st, (result1))

    # and a few very simple cases, with no contracted indices:

    t = G(m0)
    st = tfunc(t)
    assert _is_tensor_eq(st, t)

    t = -7*G(m0)
    st = tfunc(t)
    assert _is_tensor_eq(st, t)

    t = 224*G(m0)*G(m1)*G(-m2)*G(m3)
    st = tfunc(t)
    assert _is_tensor_eq(st, t)


def test_kahane_algorithm():
    # Wrap this function to convert to and from TIDS:

    def tfunc(e):
        return _simplify_single_line(e)

    execute_gamma_simplify_tests_for_function(tfunc, D=4)


def test_kahane_simplify1():
    i0,i1,i2,i3,i4,i5,i6,i7,i8,i9,i10,i11,i12,i13,i14,i15 = tensor_indices('i0:16', LorentzIndex)
    mu, nu, rho, sigma = tensor_indices("mu, nu, rho, sigma", LorentzIndex)
    D = 4
    t = G(i0)*G(i1)
    r = kahane_simplify(t)
    assert r.equals(t)

    t = G(i0)*G(i1)*G(-i0)
    r = kahane_simplify(t)
    assert r.equals(-2*G(i1))
    t = G(i0)*G(i1)*G(-i0)
    r = kahane_simplify(t)
    assert r.equals(-2*G(i1))

    t = G(i0)*G(i1)
    r = kahane_simplify(t)
    assert r.equals(t)
    t = G(i0)*G(i1)
    r = kahane_simplify(t)
    assert r.equals(t)
    t = G(i0)*G(-i0)
    r = kahane_simplify(t)
    assert r.equals(4*eye(4))
    t = G(i0)*G(-i0)
    r = kahane_simplify(t)
    assert r.equals(4*eye(4))
    t = G(i0)*G(-i0)
    r = kahane_simplify(t)
    assert r.equals(4*eye(4))
    t = G(i0)*G(i1)*G(-i0)
    r = kahane_simplify(t)
    assert r.equals(-2*G(i1))
    t = G(i0)*G(i1)*G(-i0)*G(-i1)
    r = kahane_simplify(t)
    assert r.equals((2*D - D**2)*eye(4))
    t = G(i0)*G(i1)*G(-i0)*G(-i1)
    r = kahane_simplify(t)
    assert r.equals((2*D - D**2)*eye(4))
    t = G(i0)*G(-i0)*G(i1)*G(-i1)
    r = kahane_simplify(t)
    assert r.equals(16*eye(4))
    t = (G(mu)*G(nu)*G(-nu)*G(-mu))
    r = kahane_simplify(t)
    assert r.equals(D**2*eye(4))
    t = (G(mu)*G(nu)*G(-nu)*G(-mu))
    r = kahane_simplify(t)
    assert r.equals(D**2*eye(4))
    t = (G(mu)*G(nu)*G(-nu)*G(-mu))
    r = kahane_simplify(t)
    assert r.equals(D**2*eye(4))
    t = (G(mu)*G(nu)*G(-rho)*G(-nu)*G(-mu)*G(rho))
    r = kahane_simplify(t)
    assert r.equals((4*D - 4*D**2 + D**3)*eye(4))
    t = (G(-mu)*G(-nu)*G(-rho)*G(-sigma)*G(nu)*G(mu)*G(sigma)*G(rho))
    r = kahane_simplify(t)
    assert r.equals((-16*D + 24*D**2 - 8*D**3 + D**4)*eye(4))
    t = (G(-mu)*G(nu)*G(-rho)*G(sigma)*G(rho)*G(-nu)*G(mu)*G(-sigma))
    r = kahane_simplify(t)
    assert r.equals((8*D - 12*D**2 + 6*D**3 - D**4)*eye(4))

    # Expressions with free indices:
    t = (G(mu)*G(nu)*G(rho)*G(sigma)*G(-mu))
    r = kahane_simplify(t)
    assert r.equals(-2*G(sigma)*G(rho)*G(nu))
    t = (G(mu)*G(-mu)*G(rho)*G(sigma))
    r = kahane_simplify(t)
    assert r.equals(4*G(rho)*G(sigma))
    t = (G(rho)*G(sigma)*G(mu)*G(-mu))
    r = kahane_simplify(t)
    assert r.equals(4*G(rho)*G(sigma))

def test_gamma_matrix_class():
    i, j, k = tensor_indices('i,j,k', LorentzIndex)

    # define another type of TensorHead to see if exprs are correctly handled:
    A = TensorHead('A', [LorentzIndex])

    t = A(k)*G(i)*G(-i)
    ts = simplify_gamma_expression(t)
    assert _is_tensor_eq(ts, Matrix([
        [4, 0, 0, 0],
        [0, 4, 0, 0],
        [0, 0, 4, 0],
        [0, 0, 0, 4]])*A(k))

    t = G(i)*A(k)*G(j)
    ts = simplify_gamma_expression(t)
    assert _is_tensor_eq(ts, A(k)*G(i)*G(j))

    execute_gamma_simplify_tests_for_function(simplify_gamma_expression, D=4)


def test_gamma_matrix_trace():
    g = LorentzIndex.metric

    m0, m1, m2, m3, m4, m5, m6 = tensor_indices('m0:7', LorentzIndex)
    n0, n1, n2, n3, n4, n5 = tensor_indices('n0:6', LorentzIndex)

    # working in D=4 dimensions
    D = 4

    # traces of odd number of gamma matrices are zero:
    t = G(m0)
    t1 = gamma_trace(t)
    assert t1.equals(0)

    t = G(m0)*G(m1)*G(m2)
    t1 = gamma_trace(t)
    assert t1.equals(0)

    t = G(m0)*G(m1)*G(-m0)
    t1 = gamma_trace(t)
    assert t1.equals(0)

    t = G(m0)*G(m1)*G(m2)*G(m3)*G(m4)
    t1 = gamma_trace(t)
    assert t1.equals(0)

    # traces without internal contractions:
    t = G(m0)*G(m1)
    t1 = gamma_trace(t)
    assert _is_tensor_eq(t1, 4*g(m0, m1))

    t = G(m0)*G(m1)*G(m2)*G(m3)
    t1 = gamma_trace(t)
    t2 = -4*g(m0, m2)*g(m1, m3) + 4*g(m0, m1)*g(m2, m3) + 4*g(m0, m3)*g(m1, m2)
    assert _is_tensor_eq(t1, t2)

    t = G(m0)*G(m1)*G(m2)*G(m3)*G(m4)*G(m5)
    t1 = gamma_trace(t)
    t2 = t1*g(-m0, -m5)
    t2 = t2.contract_metric(g)
    assert _is_tensor_eq(t2, D*gamma_trace(G(m1)*G(m2)*G(m3)*G(m4)))

    # traces of expressions with internal contractions:
    t = G(m0)*G(-m0)
    t1 = gamma_trace(t)
    assert t1.equals(4*D)

    t = G(m0)*G(m1)*G(-m0)*G(-m1)
    t1 = gamma_trace(t)
    assert t1.equals(8*D - 4*D**2)

    t = G(m0)*G(m1)*G(m2)*G(m3)*G(m4)*G(-m0)
    t1 = gamma_trace(t)
    t2 = (-4*D)*g(m1, m3)*g(m2, m4) + (4*D)*g(m1, m2)*g(m3, m4) + \
                 (4*D)*g(m1, m4)*g(m2, m3)
    assert _is_tensor_eq(t1, t2)

    t = G(-m5)*G(m0)*G(m1)*G(m2)*G(m3)*G(m4)*G(-m0)*G(m5)
    t1 = gamma_trace(t)
    t2 = (32*D + 4*(-D + 4)**2 - 64)*(g(m1, m2)*g(m3, m4) - \
            g(m1, m3)*g(m2, m4) + g(m1, m4)*g(m2, m3))
    assert _is_tensor_eq(t1, t2)

    t = G(m0)*G(m1)*G(-m0)*G(m3)
    t1 = gamma_trace(t)
    assert t1.equals((-4*D + 8)*g(m1, m3))

#    p, q = S1('p,q')
#    ps = p(m0)*G(-m0)
#    qs = q(m0)*G(-m0)
#    t = ps*qs*ps*qs
#    t1 = gamma_trace(t)
#    assert t1 == 8*p(m0)*q(-m0)*p(m1)*q(-m1) - 4*p(m0)*p(-m0)*q(m1)*q(-m1)

    t = G(m0)*G(m1)*G(m2)*G(m3)*G(m4)*G(m5)*G(-m0)*G(-m1)*G(-m2)*G(-m3)*G(-m4)*G(-m5)
    t1 = gamma_trace(t)
    assert t1.equals(-4*D**6 + 120*D**5 - 1040*D**4 + 3360*D**3 - 4480*D**2 + 2048*D)

    t = G(m0)*G(m1)*G(n1)*G(m2)*G(n2)*G(m3)*G(m4)*G(-n2)*G(-n1)*G(-m0)*G(-m1)*G(-m2)*G(-m3)*G(-m4)
    t1 = gamma_trace(t)
    tresu = -7168*D + 16768*D**2 - 14400*D**3 + 5920*D**4 - 1232*D**5 + 120*D**6 - 4*D**7
    assert t1.equals(tresu)

    # checked with Mathematica
    # In[1]:= <<Tracer.m
    # In[2]:= Spur[l];
    # In[3]:= GammaTrace[l, {m0},{m1},{n1},{m2},{n2},{m3},{m4},{n3},{n4},{m0},{m1},{m2},{m3},{m4}]
    t = G(m0)*G(m1)*G(n1)*G(m2)*G(n2)*G(m3)*G(m4)*G(n3)*G(n4)*G(-m0)*G(-m1)*G(-m2)*G(-m3)*G(-m4)
    t1 = gamma_trace(t)
#    t1 = t1.expand_coeff()
    c1 = -4*D**5 + 120*D**4 - 1200*D**3 + 5280*D**2 - 10560*D + 7808
    c2 = -4*D**5 + 88*D**4 - 560*D**3 + 1440*D**2 - 1600*D + 640
    assert _is_tensor_eq(t1, c1*g(n1, n4)*g(n2, n3) + c2*g(n1, n2)*g(n3, n4) + \
            (-c1)*g(n1, n3)*g(n2, n4))

    p, q = tensor_heads('p,q', [LorentzIndex])
    ps = p(m0)*G(-m0)
    qs = q(m0)*G(-m0)
    p2 = p(m0)*p(-m0)
    q2 = q(m0)*q(-m0)
    pq = p(m0)*q(-m0)
    t = ps*qs*ps*qs
    r = gamma_trace(t)
    assert _is_tensor_eq(r, 8*pq*pq - 4*p2*q2)
    t = ps*qs*ps*qs*ps*qs
    r = gamma_trace(t)
    assert _is_tensor_eq(r, -12*p2*pq*q2 + 16*pq*pq*pq)
    t = ps*qs*ps*qs*ps*qs*ps*qs
    r = gamma_trace(t)
    assert _is_tensor_eq(r, -32*pq*pq*p2*q2 + 32*pq*pq*pq*pq + 4*p2*p2*q2*q2)

    t = 4*p(m1)*p(m0)*p(-m0)*q(-m1)*q(m2)*q(-m2)
    assert _is_tensor_eq(gamma_trace(t), t)
    t = ps*ps*ps*ps*ps*ps*ps*ps
    r = gamma_trace(t)
    assert r.equals(4*p2*p2*p2*p2)


def test_bug_13636():
    """Test issue 13636 regarding handling traces of sums of products
    of GammaMatrix mixed with other factors."""
    pi, ki, pf = tensor_heads("pi, ki, pf", [LorentzIndex])
    i0, i1, i2, i3, i4 = tensor_indices("i0:5", LorentzIndex)
    x = Symbol("x")
    pis = pi(i2) * G(-i2)
    kis = ki(i3) * G(-i3)
    pfs = pf(i4) * G(-i4)

    a = pfs * G(i0) * kis * G(i1) * pis * G(-i1) * kis * G(-i0)
    b = pfs * G(i0) * kis * G(i1) * pis * x * G(-i0) * pi(-i1)
    ta = gamma_trace(a)
    tb = gamma_trace(b)
    t_a_plus_b = gamma_trace(a + b)
    assert ta == 4 * (
        -4 * ki(i0) * ki(-i0) * pf(i1) * pi(-i1)
        + 8 * ki(i0) * ki(i1) * pf(-i0) * pi(-i1)
    )
    assert tb == -8 * x * ki(i0) * pf(-i0) * pi(i1) * pi(-i1)
    assert t_a_plus_b == ta + tb

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\extension\test_numpy.py ===
"""
This file contains a minimal set of tests for compliance with the extension
array interface test suite, and should contain no other tests.
The test suite for the full functionality of the array is located in
`pandas/tests/arrays/`.

The tests in this file are inherited from the BaseExtensionTests, and only
minimal tweaks should be applied to get the tests passing (by overwriting a
parent method).

Additional tests should either be added to one of the BaseExtensionTests
classes (if they are relevant for the extension interface for all dtypes), or
be added to the array-specific tests in `pandas/tests/arrays/`.

Note: we do not bother with base.BaseIndexTests because NumpyExtensionArray
will never be held in an Index.
"""
import numpy as np
import pytest

from pandas.core.dtypes.dtypes import NumpyEADtype

import pandas as pd
import pandas._testing as tm
from pandas.api.types import is_object_dtype
from pandas.core.arrays.numpy_ import NumpyExtensionArray
from pandas.tests.extension import base

orig_assert_attr_equal = tm.assert_attr_equal


def _assert_attr_equal(attr: str, left, right, obj: str = "Attributes"):
    """
    patch tm.assert_attr_equal so NumpyEADtype("object") is closed enough to
    np.dtype("object")
    """
    if attr == "dtype":
        lattr = getattr(left, "dtype", None)
        rattr = getattr(right, "dtype", None)
        if isinstance(lattr, NumpyEADtype) and not isinstance(rattr, NumpyEADtype):
            left = left.astype(lattr.numpy_dtype)
        elif isinstance(rattr, NumpyEADtype) and not isinstance(lattr, NumpyEADtype):
            right = right.astype(rattr.numpy_dtype)

    orig_assert_attr_equal(attr, left, right, obj)


@pytest.fixture(params=["float", "object"])
def dtype(request):
    return NumpyEADtype(np.dtype(request.param))


@pytest.fixture
def allow_in_pandas(monkeypatch):
    """
    A monkeypatch to tells pandas to let us in.

    By default, passing a NumpyExtensionArray to an index / series / frame
    constructor will unbox that NumpyExtensionArray to an ndarray, and treat
    it as a non-EA column. We don't want people using EAs without
    reason.

    The mechanism for this is a check against ABCNumpyExtensionArray
    in each constructor.

    But, for testing, we need to allow them in pandas. So we patch
    the _typ of NumpyExtensionArray, so that we evade the ABCNumpyExtensionArray
    check.
    """
    with monkeypatch.context() as m:
        m.setattr(NumpyExtensionArray, "_typ", "extension")
        m.setattr(tm.asserters, "assert_attr_equal", _assert_attr_equal)
        yield


@pytest.fixture
def data(allow_in_pandas, dtype):
    if dtype.numpy_dtype == "object":
        return pd.Series([(i,) for i in range(100)]).array
    return NumpyExtensionArray(np.arange(1, 101, dtype=dtype._dtype))


@pytest.fixture
def data_missing(allow_in_pandas, dtype):
    if dtype.numpy_dtype == "object":
        return NumpyExtensionArray(np.array([np.nan, (1,)], dtype=object))
    return NumpyExtensionArray(np.array([np.nan, 1.0]))


@pytest.fixture
def na_cmp():
    def cmp(a, b):
        return np.isnan(a) and np.isnan(b)

    return cmp


@pytest.fixture
def data_for_sorting(allow_in_pandas, dtype):
    """Length-3 array with a known sort order.

    This should be three items [B, C, A] with
    A < B < C
    """
    if dtype.numpy_dtype == "object":
        # Use an empty tuple for first element, then remove,
        # to disable np.array's shape inference.
        return NumpyExtensionArray(np.array([(), (2,), (3,), (1,)], dtype=object)[1:])
    return NumpyExtensionArray(np.array([1, 2, 0]))


@pytest.fixture
def data_missing_for_sorting(allow_in_pandas, dtype):
    """Length-3 array with a known sort order.

    This should be three items [B, NA, A] with
    A < B and NA missing.
    """
    if dtype.numpy_dtype == "object":
        return NumpyExtensionArray(np.array([(1,), np.nan, (0,)], dtype=object))
    return NumpyExtensionArray(np.array([1, np.nan, 0]))


@pytest.fixture
def data_for_grouping(allow_in_pandas, dtype):
    """Data for factorization, grouping, and unique tests.

    Expected to be like [B, B, NA, NA, A, A, B, C]

    Where A < B < C and NA is missing
    """
    if dtype.numpy_dtype == "object":
        a, b, c = (1,), (2,), (3,)
    else:
        a, b, c = np.arange(3)
    return NumpyExtensionArray(
        np.array([b, b, np.nan, np.nan, a, a, b, c], dtype=dtype.numpy_dtype)
    )


@pytest.fixture
def data_for_twos(dtype):
    if dtype.kind == "O":
        pytest.skip(f"{dtype} is not a numeric dtype")
    arr = np.ones(100) * 2
    return NumpyExtensionArray._from_sequence(arr, dtype=dtype)


@pytest.fixture
def skip_numpy_object(dtype, request):
    """
    Tests for NumpyExtensionArray with nested data. Users typically won't create
    these objects via `pd.array`, but they can show up through `.array`
    on a Series with nested data. Many of the base tests fail, as they aren't
    appropriate for nested data.

    This fixture allows these tests to be skipped when used as a usefixtures
    marker to either an individual test or a test class.
    """
    if dtype == "object":
        mark = pytest.mark.xfail(reason="Fails for object dtype")
        request.applymarker(mark)


skip_nested = pytest.mark.usefixtures("skip_numpy_object")


class TestNumpyExtensionArray(base.ExtensionTests):
    @pytest.mark.skip(reason="We don't register our dtype")
    # We don't want to register. This test should probably be split in two.
    def test_from_dtype(self, data):
        pass

    @skip_nested
    def test_series_constructor_scalar_with_index(self, data, dtype):
        # ValueError: Length of passed values is 1, index implies 3.
        super().test_series_constructor_scalar_with_index(data, dtype)

    def test_check_dtype(self, data, request, using_infer_string):
        if data.dtype.numpy_dtype == "object":
            request.applymarker(
                pytest.mark.xfail(
                    reason=f"NumpyExtensionArray expectedly clashes with a "
                    f"NumPy name: {data.dtype.numpy_dtype}"
                )
            )
        super().test_check_dtype(data)

    def test_is_not_object_type(self, dtype, request):
        if dtype.numpy_dtype == "object":
            # Different from BaseDtypeTests.test_is_not_object_type
            # because NumpyEADtype(object) is an object type
            assert is_object_dtype(dtype)
        else:
            super().test_is_not_object_type(dtype)

    @skip_nested
    def test_getitem_scalar(self, data):
        # AssertionError
        super().test_getitem_scalar(data)

    @skip_nested
    def test_shift_fill_value(self, data):
        # np.array shape inference. Shift implementation fails.
        super().test_shift_fill_value(data)

    @skip_nested
    def test_fillna_copy_frame(self, data_missing):
        # The "scalar" for this array isn't a scalar.
        super().test_fillna_copy_frame(data_missing)

    @skip_nested
    def test_fillna_copy_series(self, data_missing):
        # The "scalar" for this array isn't a scalar.
        super().test_fillna_copy_series(data_missing)

    @skip_nested
    def test_searchsorted(self, data_for_sorting, as_series):
        # TODO: NumpyExtensionArray.searchsorted calls ndarray.searchsorted which
        #  isn't quite what we want in nested data cases. Instead we need to
        #  adapt something like libindex._bin_search.
        super().test_searchsorted(data_for_sorting, as_series)

    @pytest.mark.xfail(reason="NumpyExtensionArray.diff may fail on dtype")
    def test_diff(self, data, periods):
        return super().test_diff(data, periods)

    def test_insert(self, data, request):
        if data.dtype.numpy_dtype == object:
            mark = pytest.mark.xfail(reason="Dimension mismatch in np.concatenate")
            request.applymarker(mark)

        super().test_insert(data)

    @skip_nested
    def test_insert_invalid(self, data, invalid_scalar):
        # NumpyExtensionArray[object] can hold anything, so skip
        super().test_insert_invalid(data, invalid_scalar)

    divmod_exc = None
    series_scalar_exc = None
    frame_scalar_exc = None
    series_array_exc = None

    def test_divmod(self, data):
        divmod_exc = None
        if data.dtype.kind == "O":
            divmod_exc = TypeError
        self.divmod_exc = divmod_exc
        super().test_divmod(data)

    def test_divmod_series_array(self, data):
        ser = pd.Series(data)
        exc = None
        if data.dtype.kind == "O":
            exc = TypeError
            self.divmod_exc = exc
        self._check_divmod_op(ser, divmod, data)

    def test_arith_series_with_scalar(self, data, all_arithmetic_operators, request):
        opname = all_arithmetic_operators
        series_scalar_exc = None
        if data.dtype.numpy_dtype == object:
            if opname in ["__mul__", "__rmul__"]:
                mark = pytest.mark.xfail(
                    reason="the Series.combine step raises but not the Series method."
                )
                request.node.add_marker(mark)
            series_scalar_exc = TypeError
        self.series_scalar_exc = series_scalar_exc
        super().test_arith_series_with_scalar(data, all_arithmetic_operators)

    def test_arith_series_with_array(self, data, all_arithmetic_operators):
        opname = all_arithmetic_operators
        series_array_exc = None
        if data.dtype.numpy_dtype == object and opname not in ["__add__", "__radd__"]:
            series_array_exc = TypeError
        self.series_array_exc = series_array_exc
        super().test_arith_series_with_array(data, all_arithmetic_operators)

    def test_arith_frame_with_scalar(self, data, all_arithmetic_operators, request):
        opname = all_arithmetic_operators
        frame_scalar_exc = None
        if data.dtype.numpy_dtype == object:
            if opname in ["__mul__", "__rmul__"]:
                mark = pytest.mark.xfail(
                    reason="the Series.combine step raises but not the Series method."
                )
                request.node.add_marker(mark)
            frame_scalar_exc = TypeError
        self.frame_scalar_exc = frame_scalar_exc
        super().test_arith_frame_with_scalar(data, all_arithmetic_operators)

    def _supports_reduction(self, ser: pd.Series, op_name: str) -> bool:
        if ser.dtype.kind == "O":
            return op_name in ["sum", "min", "max", "any", "all"]
        return True

    def check_reduce(self, ser: pd.Series, op_name: str, skipna: bool):
        res_op = getattr(ser, op_name)
        # avoid coercing int -> float. Just cast to the actual numpy type.
        # error: Item "ExtensionDtype" of "dtype[Any] | ExtensionDtype" has
        # no attribute "numpy_dtype"
        cmp_dtype = ser.dtype.numpy_dtype  # type: ignore[union-attr]
        alt = ser.astype(cmp_dtype)
        exp_op = getattr(alt, op_name)
        if op_name == "count":
            result = res_op()
            expected = exp_op()
        else:
            result = res_op(skipna=skipna)
            expected = exp_op(skipna=skipna)
        tm.assert_almost_equal(result, expected)

    @pytest.mark.skip("TODO: tests not written yet")
    @pytest.mark.parametrize("skipna", [True, False])
    def test_reduce_frame(self, data, all_numeric_reductions, skipna):
        pass

    @skip_nested
    def test_fillna_series(self, data_missing):
        # Non-scalar "scalar" values.
        super().test_fillna_series(data_missing)

    @skip_nested
    def test_fillna_frame(self, data_missing):
        # Non-scalar "scalar" values.
        super().test_fillna_frame(data_missing)

    @skip_nested
    def test_setitem_invalid(self, data, invalid_scalar):
        # object dtype can hold anything, so doesn't raise
        super().test_setitem_invalid(data, invalid_scalar)

    @skip_nested
    def test_setitem_sequence_broadcasts(self, data, box_in_series):
        # ValueError: cannot set using a list-like indexer with a different
        # length than the value
        super().test_setitem_sequence_broadcasts(data, box_in_series)

    @skip_nested
    @pytest.mark.parametrize("setter", ["loc", None])
    def test_setitem_mask_broadcast(self, data, setter):
        # ValueError: cannot set using a list-like indexer with a different
        # length than the value
        super().test_setitem_mask_broadcast(data, setter)

    @skip_nested
    def test_setitem_scalar_key_sequence_raise(self, data):
        # Failed: DID NOT RAISE <class 'ValueError'>
        super().test_setitem_scalar_key_sequence_raise(data)

    # TODO: there is some issue with NumpyExtensionArray, therefore,
    #   skip the setitem test for now, and fix it later (GH 31446)

    @skip_nested
    @pytest.mark.parametrize(
        "mask",
        [
            np.array([True, True, True, False, False]),
            pd.array([True, True, True, False, False], dtype="boolean"),
        ],
        ids=["numpy-array", "boolean-array"],
    )
    def test_setitem_mask(self, data, mask, box_in_series):
        super().test_setitem_mask(data, mask, box_in_series)

    @skip_nested
    @pytest.mark.parametrize(
        "idx",
        [[0, 1, 2], pd.array([0, 1, 2], dtype="Int64"), np.array([0, 1, 2])],
        ids=["list", "integer-array", "numpy-array"],
    )
    def test_setitem_integer_array(self, data, idx, box_in_series):
        super().test_setitem_integer_array(data, idx, box_in_series)

    @pytest.mark.parametrize(
        "idx, box_in_series",
        [
            ([0, 1, 2, pd.NA], False),
            pytest.param([0, 1, 2, pd.NA], True, marks=pytest.mark.xfail),
            (pd.array([0, 1, 2, pd.NA], dtype="Int64"), False),
            (pd.array([0, 1, 2, pd.NA], dtype="Int64"), False),
        ],
        ids=["list-False", "list-True", "integer-array-False", "integer-array-True"],
    )
    def test_setitem_integer_with_missing_raises(self, data, idx, box_in_series):
        super().test_setitem_integer_with_missing_raises(data, idx, box_in_series)

    @skip_nested
    def test_setitem_slice(self, data, box_in_series):
        super().test_setitem_slice(data, box_in_series)

    @skip_nested
    def test_setitem_loc_iloc_slice(self, data):
        super().test_setitem_loc_iloc_slice(data)

    def test_setitem_with_expansion_dataframe_column(self, data, full_indexer):
        # https://github.com/pandas-dev/pandas/issues/32395
        df = expected = pd.DataFrame({"data": pd.Series(data)})
        result = pd.DataFrame(index=df.index)

        # because result has object dtype, the attempt to do setting inplace
        #  is successful, and object dtype is retained
        key = full_indexer(df)
        result.loc[key, "data"] = df["data"]

        # base class method has expected = df; NumpyExtensionArray behaves oddly because
        #  we patch _typ for these tests.
        if data.dtype.numpy_dtype != object:
            if not isinstance(key, slice) or key != slice(None):
                expected = pd.DataFrame({"data": data.to_numpy()})
        tm.assert_frame_equal(result, expected, check_column_type=False)

    @pytest.mark.xfail(reason="NumpyEADtype is unpacked")
    def test_index_from_listlike_with_dtype(self, data):
        super().test_index_from_listlike_with_dtype(data)

    @skip_nested
    @pytest.mark.parametrize("engine", ["c", "python"])
    def test_EA_types(self, engine, data, request):
        super().test_EA_types(engine, data, request)


class Test2DCompat(base.NDArrayBacked2DTests):
    pass

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\strings\test_case_justify.py ===
from datetime import datetime
import operator

import numpy as np
import pytest

from pandas import (
    Series,
    _testing as tm,
)


def test_title(any_string_dtype):
    s = Series(["FOO", "BAR", np.nan, "Blah", "blurg"], dtype=any_string_dtype)
    result = s.str.title()
    expected = Series(["Foo", "Bar", np.nan, "Blah", "Blurg"], dtype=any_string_dtype)
    tm.assert_series_equal(result, expected)


def test_title_mixed_object():
    s = Series(["FOO", np.nan, "bar", True, datetime.today(), "blah", None, 1, 2.0])
    result = s.str.title()
    expected = Series(
        ["Foo", np.nan, "Bar", np.nan, np.nan, "Blah", None, np.nan, np.nan],
        dtype=object,
    )
    tm.assert_almost_equal(result, expected)


def test_lower_upper(any_string_dtype):
    s = Series(["om", np.nan, "nom", "nom"], dtype=any_string_dtype)

    result = s.str.upper()
    expected = Series(["OM", np.nan, "NOM", "NOM"], dtype=any_string_dtype)
    tm.assert_series_equal(result, expected)

    result = result.str.lower()
    tm.assert_series_equal(result, s)


def test_lower_upper_mixed_object():
    s = Series(["a", np.nan, "b", True, datetime.today(), "foo", None, 1, 2.0])

    result = s.str.upper()
    expected = Series(
        ["A", np.nan, "B", np.nan, np.nan, "FOO", None, np.nan, np.nan], dtype=object
    )
    tm.assert_series_equal(result, expected)

    result = s.str.lower()
    expected = Series(
        ["a", np.nan, "b", np.nan, np.nan, "foo", None, np.nan, np.nan], dtype=object
    )
    tm.assert_series_equal(result, expected)


@pytest.mark.parametrize(
    "data, expected",
    [
        (
            ["FOO", "BAR", np.nan, "Blah", "blurg"],
            ["Foo", "Bar", np.nan, "Blah", "Blurg"],
        ),
        (["a", "b", "c"], ["A", "B", "C"]),
        (["a b", "a bc. de"], ["A b", "A bc. de"]),
    ],
)
def test_capitalize(data, expected, any_string_dtype):
    s = Series(data, dtype=any_string_dtype)
    result = s.str.capitalize()
    expected = Series(expected, dtype=any_string_dtype)
    tm.assert_series_equal(result, expected)


def test_capitalize_mixed_object():
    s = Series(["FOO", np.nan, "bar", True, datetime.today(), "blah", None, 1, 2.0])
    result = s.str.capitalize()
    expected = Series(
        ["Foo", np.nan, "Bar", np.nan, np.nan, "Blah", None, np.nan, np.nan],
        dtype=object,
    )
    tm.assert_series_equal(result, expected)


def test_swapcase(any_string_dtype):
    s = Series(["FOO", "BAR", np.nan, "Blah", "blurg"], dtype=any_string_dtype)
    result = s.str.swapcase()
    expected = Series(["foo", "bar", np.nan, "bLAH", "BLURG"], dtype=any_string_dtype)
    tm.assert_series_equal(result, expected)


def test_swapcase_mixed_object():
    s = Series(["FOO", np.nan, "bar", True, datetime.today(), "Blah", None, 1, 2.0])
    result = s.str.swapcase()
    expected = Series(
        ["foo", np.nan, "BAR", np.nan, np.nan, "bLAH", None, np.nan, np.nan],
        dtype=object,
    )
    tm.assert_series_equal(result, expected)


def test_casefold():
    # GH25405
    expected = Series(["ss", np.nan, "case", "ssd"])
    s = Series(["ß", np.nan, "case", "ßd"])
    result = s.str.casefold()

    tm.assert_series_equal(result, expected)


def test_casemethods(any_string_dtype):
    values = ["aaa", "bbb", "CCC", "Dddd", "eEEE"]
    s = Series(values, dtype=any_string_dtype)
    assert s.str.lower().tolist() == [v.lower() for v in values]
    assert s.str.upper().tolist() == [v.upper() for v in values]
    assert s.str.title().tolist() == [v.title() for v in values]
    assert s.str.capitalize().tolist() == [v.capitalize() for v in values]
    assert s.str.swapcase().tolist() == [v.swapcase() for v in values]


def test_pad(any_string_dtype):
    s = Series(["a", "b", np.nan, "c", np.nan, "eeeeee"], dtype=any_string_dtype)

    result = s.str.pad(5, side="left")
    expected = Series(
        ["    a", "    b", np.nan, "    c", np.nan, "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)

    result = s.str.pad(5, side="right")
    expected = Series(
        ["a    ", "b    ", np.nan, "c    ", np.nan, "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)

    result = s.str.pad(5, side="both")
    expected = Series(
        ["  a  ", "  b  ", np.nan, "  c  ", np.nan, "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)


def test_pad_mixed_object():
    s = Series(["a", np.nan, "b", True, datetime.today(), "ee", None, 1, 2.0])

    result = s.str.pad(5, side="left")
    expected = Series(
        ["    a", np.nan, "    b", np.nan, np.nan, "   ee", None, np.nan, np.nan],
        dtype=object,
    )
    tm.assert_series_equal(result, expected)

    result = s.str.pad(5, side="right")
    expected = Series(
        ["a    ", np.nan, "b    ", np.nan, np.nan, "ee   ", None, np.nan, np.nan],
        dtype=object,
    )
    tm.assert_series_equal(result, expected)

    result = s.str.pad(5, side="both")
    expected = Series(
        ["  a  ", np.nan, "  b  ", np.nan, np.nan, "  ee ", None, np.nan, np.nan],
        dtype=object,
    )
    tm.assert_series_equal(result, expected)


def test_pad_fillchar(any_string_dtype):
    s = Series(["a", "b", np.nan, "c", np.nan, "eeeeee"], dtype=any_string_dtype)

    result = s.str.pad(5, side="left", fillchar="X")
    expected = Series(
        ["XXXXa", "XXXXb", np.nan, "XXXXc", np.nan, "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)

    result = s.str.pad(5, side="right", fillchar="X")
    expected = Series(
        ["aXXXX", "bXXXX", np.nan, "cXXXX", np.nan, "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)

    result = s.str.pad(5, side="both", fillchar="X")
    expected = Series(
        ["XXaXX", "XXbXX", np.nan, "XXcXX", np.nan, "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)


def test_pad_fillchar_bad_arg_raises(any_string_dtype):
    s = Series(["a", "b", np.nan, "c", np.nan, "eeeeee"], dtype=any_string_dtype)

    msg = "fillchar must be a character, not str"
    with pytest.raises(TypeError, match=msg):
        s.str.pad(5, fillchar="XY")

    msg = "fillchar must be a character, not int"
    with pytest.raises(TypeError, match=msg):
        s.str.pad(5, fillchar=5)


@pytest.mark.parametrize("method_name", ["center", "ljust", "rjust", "zfill", "pad"])
def test_pad_width_bad_arg_raises(method_name, any_string_dtype):
    # see gh-13598
    s = Series(["1", "22", "a", "bb"], dtype=any_string_dtype)
    op = operator.methodcaller(method_name, "f")

    msg = "width must be of integer type, not str"
    with pytest.raises(TypeError, match=msg):
        op(s.str)


def test_center_ljust_rjust(any_string_dtype):
    s = Series(["a", "b", np.nan, "c", np.nan, "eeeeee"], dtype=any_string_dtype)

    result = s.str.center(5)
    expected = Series(
        ["  a  ", "  b  ", np.nan, "  c  ", np.nan, "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)

    result = s.str.ljust(5)
    expected = Series(
        ["a    ", "b    ", np.nan, "c    ", np.nan, "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)

    result = s.str.rjust(5)
    expected = Series(
        ["    a", "    b", np.nan, "    c", np.nan, "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)


def test_center_ljust_rjust_mixed_object():
    s = Series(["a", np.nan, "b", True, datetime.today(), "c", "eee", None, 1, 2.0])

    result = s.str.center(5)
    expected = Series(
        [
            "  a  ",
            np.nan,
            "  b  ",
            np.nan,
            np.nan,
            "  c  ",
            " eee ",
            None,
            np.nan,
            np.nan,
        ],
        dtype=object,
    )
    tm.assert_series_equal(result, expected)

    result = s.str.ljust(5)
    expected = Series(
        [
            "a    ",
            np.nan,
            "b    ",
            np.nan,
            np.nan,
            "c    ",
            "eee  ",
            None,
            np.nan,
            np.nan,
        ],
        dtype=object,
    )
    tm.assert_series_equal(result, expected)

    result = s.str.rjust(5)
    expected = Series(
        [
            "    a",
            np.nan,
            "    b",
            np.nan,
            np.nan,
            "    c",
            "  eee",
            None,
            np.nan,
            np.nan,
        ],
        dtype=object,
    )
    tm.assert_series_equal(result, expected)


def test_center_ljust_rjust_fillchar(any_string_dtype):
    # GH#54533, GH#54792
    s = Series(["a", "bb", "cccc", "ddddd", "eeeeee"], dtype=any_string_dtype)

    result = s.str.center(5, fillchar="X")
    expected = Series(
        ["XXaXX", "XXbbX", "Xcccc", "ddddd", "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)
    expected = np.array([v.center(5, "X") for v in np.array(s)], dtype=np.object_)
    tm.assert_numpy_array_equal(np.array(result, dtype=np.object_), expected)

    result = s.str.ljust(5, fillchar="X")
    expected = Series(
        ["aXXXX", "bbXXX", "ccccX", "ddddd", "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)
    expected = np.array([v.ljust(5, "X") for v in np.array(s)], dtype=np.object_)
    tm.assert_numpy_array_equal(np.array(result, dtype=np.object_), expected)

    result = s.str.rjust(5, fillchar="X")
    expected = Series(
        ["XXXXa", "XXXbb", "Xcccc", "ddddd", "eeeeee"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)
    expected = np.array([v.rjust(5, "X") for v in np.array(s)], dtype=np.object_)
    tm.assert_numpy_array_equal(np.array(result, dtype=np.object_), expected)


def test_center_ljust_rjust_fillchar_bad_arg_raises(any_string_dtype):
    s = Series(["a", "bb", "cccc", "ddddd", "eeeeee"], dtype=any_string_dtype)

    # If fillchar is not a character, normal str raises TypeError
    # 'aaa'.ljust(5, 'XY')
    # TypeError: must be char, not str
    template = "fillchar must be a character, not {dtype}"

    with pytest.raises(TypeError, match=template.format(dtype="str")):
        s.str.center(5, fillchar="XY")

    with pytest.raises(TypeError, match=template.format(dtype="str")):
        s.str.ljust(5, fillchar="XY")

    with pytest.raises(TypeError, match=template.format(dtype="str")):
        s.str.rjust(5, fillchar="XY")

    with pytest.raises(TypeError, match=template.format(dtype="int")):
        s.str.center(5, fillchar=1)

    with pytest.raises(TypeError, match=template.format(dtype="int")):
        s.str.ljust(5, fillchar=1)

    with pytest.raises(TypeError, match=template.format(dtype="int")):
        s.str.rjust(5, fillchar=1)


def test_zfill(any_string_dtype):
    s = Series(["1", "22", "aaa", "333", "45678"], dtype=any_string_dtype)

    result = s.str.zfill(5)
    expected = Series(
        ["00001", "00022", "00aaa", "00333", "45678"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)
    expected = np.array([v.zfill(5) for v in np.array(s)], dtype=np.object_)
    tm.assert_numpy_array_equal(np.array(result, dtype=np.object_), expected)

    result = s.str.zfill(3)
    expected = Series(["001", "022", "aaa", "333", "45678"], dtype=any_string_dtype)
    tm.assert_series_equal(result, expected)
    expected = np.array([v.zfill(3) for v in np.array(s)], dtype=np.object_)
    tm.assert_numpy_array_equal(np.array(result, dtype=np.object_), expected)

    s = Series(["1", np.nan, "aaa", np.nan, "45678"], dtype=any_string_dtype)
    result = s.str.zfill(5)
    expected = Series(
        ["00001", np.nan, "00aaa", np.nan, "45678"], dtype=any_string_dtype
    )
    tm.assert_series_equal(result, expected)


def test_wrap(any_string_dtype):
    # test values are: two words less than width, two words equal to width,
    # two words greater than width, one word less than width, one word
    # equal to width, one word greater than width, multiple tokens with
    # trailing whitespace equal to width
    s = Series(
        [
            "hello world",
            "hello world!",
            "hello world!!",
            "abcdefabcde",
            "abcdefabcdef",
            "abcdefabcdefa",
            "ab ab ab ab ",
            "ab ab ab ab a",
            "\t",
        ],
        dtype=any_string_dtype,
    )

    # expected values
    expected = Series(
        [
            "hello world",
            "hello world!",
            "hello\nworld!!",
            "abcdefabcde",
            "abcdefabcdef",
            "abcdefabcdef\na",
            "ab ab ab ab",
            "ab ab ab ab\na",
            "",
        ],
        dtype=any_string_dtype,
    )

    result = s.str.wrap(12, break_long_words=True)
    tm.assert_series_equal(result, expected)


def test_wrap_unicode(any_string_dtype):
    # test with pre and post whitespace (non-unicode), NaN, and non-ascii Unicode
    s = Series(
        ["  pre  ", np.nan, "\xac\u20ac\U00008000 abadcafe"], dtype=any_string_dtype
    )
    expected = Series(
        ["  pre", np.nan, "\xac\u20ac\U00008000 ab\nadcafe"], dtype=any_string_dtype
    )
    result = s.str.wrap(6)
    tm.assert_series_equal(result, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\f2py\tests\test_crackfortran.py ===
import contextlib
import importlib
import io
import textwrap
import time

import pytest

import numpy as np
from numpy.f2py import crackfortran
from numpy.f2py.crackfortran import markinnerspaces, nameargspattern

from . import util


class TestNoSpace(util.F2PyTest):
    # issue gh-15035: add handling for endsubroutine, endfunction with no space
    # between "end" and the block name
    sources = [util.getpath("tests", "src", "crackfortran", "gh15035.f")]

    def test_module(self):
        k = np.array([1, 2, 3], dtype=np.float64)
        w = np.array([1, 2, 3], dtype=np.float64)
        self.module.subb(k)
        assert np.allclose(k, w + 1)
        self.module.subc([w, k])
        assert np.allclose(k, w + 1)
        assert self.module.t0("23") == b"2"


class TestPublicPrivate:
    def test_defaultPrivate(self):
        fpath = util.getpath("tests", "src", "crackfortran", "privatemod.f90")
        mod = crackfortran.crackfortran([str(fpath)])
        assert len(mod) == 1
        mod = mod[0]
        assert "private" in mod["vars"]["a"]["attrspec"]
        assert "public" not in mod["vars"]["a"]["attrspec"]
        assert "private" in mod["vars"]["b"]["attrspec"]
        assert "public" not in mod["vars"]["b"]["attrspec"]
        assert "private" not in mod["vars"]["seta"]["attrspec"]
        assert "public" in mod["vars"]["seta"]["attrspec"]

    def test_defaultPublic(self, tmp_path):
        fpath = util.getpath("tests", "src", "crackfortran", "publicmod.f90")
        mod = crackfortran.crackfortran([str(fpath)])
        assert len(mod) == 1
        mod = mod[0]
        assert "private" in mod["vars"]["a"]["attrspec"]
        assert "public" not in mod["vars"]["a"]["attrspec"]
        assert "private" not in mod["vars"]["seta"]["attrspec"]
        assert "public" in mod["vars"]["seta"]["attrspec"]

    def test_access_type(self, tmp_path):
        fpath = util.getpath("tests", "src", "crackfortran", "accesstype.f90")
        mod = crackfortran.crackfortran([str(fpath)])
        assert len(mod) == 1
        tt = mod[0]['vars']
        assert set(tt['a']['attrspec']) == {'private', 'bind(c)'}
        assert set(tt['b_']['attrspec']) == {'public', 'bind(c)'}
        assert set(tt['c']['attrspec']) == {'public'}

    def test_nowrap_private_proceedures(self, tmp_path):
        fpath = util.getpath("tests", "src", "crackfortran", "gh23879.f90")
        mod = crackfortran.crackfortran([str(fpath)])
        assert len(mod) == 1
        pyf = crackfortran.crack2fortran(mod)
        assert 'bar' not in pyf

class TestModuleProcedure:
    def test_moduleOperators(self, tmp_path):
        fpath = util.getpath("tests", "src", "crackfortran", "operators.f90")
        mod = crackfortran.crackfortran([str(fpath)])
        assert len(mod) == 1
        mod = mod[0]
        assert "body" in mod and len(mod["body"]) == 9
        assert mod["body"][1]["name"] == "operator(.item.)"
        assert "implementedby" in mod["body"][1]
        assert mod["body"][1]["implementedby"] == \
            ["item_int", "item_real"]
        assert mod["body"][2]["name"] == "operator(==)"
        assert "implementedby" in mod["body"][2]
        assert mod["body"][2]["implementedby"] == ["items_are_equal"]
        assert mod["body"][3]["name"] == "assignment(=)"
        assert "implementedby" in mod["body"][3]
        assert mod["body"][3]["implementedby"] == \
            ["get_int", "get_real"]

    def test_notPublicPrivate(self, tmp_path):
        fpath = util.getpath("tests", "src", "crackfortran", "pubprivmod.f90")
        mod = crackfortran.crackfortran([str(fpath)])
        assert len(mod) == 1
        mod = mod[0]
        assert mod['vars']['a']['attrspec'] == ['private', ]
        assert mod['vars']['b']['attrspec'] == ['public', ]
        assert mod['vars']['seta']['attrspec'] == ['public', ]


class TestExternal(util.F2PyTest):
    # issue gh-17859: add external attribute support
    sources = [util.getpath("tests", "src", "crackfortran", "gh17859.f")]

    def test_external_as_statement(self):
        def incr(x):
            return x + 123

        r = self.module.external_as_statement(incr)
        assert r == 123

    def test_external_as_attribute(self):
        def incr(x):
            return x + 123

        r = self.module.external_as_attribute(incr)
        assert r == 123


class TestCrackFortran(util.F2PyTest):
    # gh-2848: commented lines between parameters in subroutine parameter lists
    sources = [util.getpath("tests", "src", "crackfortran", "gh2848.f90"),
               util.getpath("tests", "src", "crackfortran", "common_with_division.f")
              ]

    def test_gh2848(self):
        r = self.module.gh2848(1, 2)
        assert r == (1, 2)

    def test_common_with_division(self):
        assert len(self.module.mortmp.ctmp) == 11

class TestMarkinnerspaces:
    # gh-14118: markinnerspaces does not handle multiple quotations

    def test_do_not_touch_normal_spaces(self):
        test_list = ["a ", " a", "a b c", "'abcdefghij'"]
        for i in test_list:
            assert markinnerspaces(i) == i

    def test_one_relevant_space(self):
        assert markinnerspaces("a 'b c' \\' \\'") == "a 'b@_@c' \\' \\'"
        assert markinnerspaces(r'a "b c" \" \"') == r'a "b@_@c" \" \"'

    def test_ignore_inner_quotes(self):
        assert markinnerspaces("a 'b c\" \" d' e") == "a 'b@_@c\"@_@\"@_@d' e"
        assert markinnerspaces("a \"b c' ' d\" e") == "a \"b@_@c'@_@'@_@d\" e"

    def test_multiple_relevant_spaces(self):
        assert markinnerspaces("a 'b c' 'd e'") == "a 'b@_@c' 'd@_@e'"
        assert markinnerspaces(r'a "b c" "d e"') == r'a "b@_@c" "d@_@e"'


class TestDimSpec(util.F2PyTest):
    """This test suite tests various expressions that are used as dimension
    specifications.

    There exists two usage cases where analyzing dimensions
    specifications are important.

    In the first case, the size of output arrays must be defined based
    on the inputs to a Fortran function. Because Fortran supports
    arbitrary bases for indexing, for instance, `arr(lower:upper)`,
    f2py has to evaluate an expression `upper - lower + 1` where
    `lower` and `upper` are arbitrary expressions of input parameters.
    The evaluation is performed in C, so f2py has to translate Fortran
    expressions to valid C expressions (an alternative approach is
    that a developer specifies the corresponding C expressions in a
    .pyf file).

    In the second case, when user provides an input array with a given
    size but some hidden parameters used in dimensions specifications
    need to be determined based on the input array size. This is a
    harder problem because f2py has to solve the inverse problem: find
    a parameter `p` such that `upper(p) - lower(p) + 1` equals to the
    size of input array. In the case when this equation cannot be
    solved (e.g. because the input array size is wrong), raise an
    error before calling the Fortran function (that otherwise would
    likely crash Python process when the size of input arrays is
    wrong). f2py currently supports this case only when the equation
    is linear with respect to unknown parameter.

    """

    suffix = ".f90"

    code_template = textwrap.dedent("""
      function get_arr_size_{count}(a, n) result (length)
        integer, intent(in) :: n
        integer, dimension({dimspec}), intent(out) :: a
        integer length
        length = size(a)
      end function

      subroutine get_inv_arr_size_{count}(a, n)
        integer :: n
        ! the value of n is computed in f2py wrapper
        !f2py intent(out) n
        integer, dimension({dimspec}), intent(in) :: a
        if (a({first}).gt.0) then
          ! print*, "a=", a
        endif
      end subroutine
    """)

    linear_dimspecs = [
        "n", "2*n", "2:n", "n/2", "5 - n/2", "3*n:20", "n*(n+1):n*(n+5)",
        "2*n, n"
    ]
    nonlinear_dimspecs = ["2*n:3*n*n+2*n"]
    all_dimspecs = linear_dimspecs + nonlinear_dimspecs

    code = ""
    for count, dimspec in enumerate(all_dimspecs):
        lst = [(d.split(":")[0] if ":" in d else "1") for d in dimspec.split(',')]
        code += code_template.format(
            count=count,
            dimspec=dimspec,
            first=", ".join(lst),
        )

    @pytest.mark.parametrize("dimspec", all_dimspecs)
    @pytest.mark.slow
    def test_array_size(self, dimspec):

        count = self.all_dimspecs.index(dimspec)
        get_arr_size = getattr(self.module, f"get_arr_size_{count}")

        for n in [1, 2, 3, 4, 5]:
            sz, a = get_arr_size(n)
            assert a.size == sz

    @pytest.mark.parametrize("dimspec", all_dimspecs)
    def test_inv_array_size(self, dimspec):

        count = self.all_dimspecs.index(dimspec)
        get_arr_size = getattr(self.module, f"get_arr_size_{count}")
        get_inv_arr_size = getattr(self.module, f"get_inv_arr_size_{count}")

        for n in [1, 2, 3, 4, 5]:
            sz, a = get_arr_size(n)
            if dimspec in self.nonlinear_dimspecs:
                # one must specify n as input, the call we'll ensure
                # that a and n are compatible:
                n1 = get_inv_arr_size(a, n)
            else:
                # in case of linear dependence, n can be determined
                # from the shape of a:
                n1 = get_inv_arr_size(a)
            # n1 may be different from n (for instance, when `a` size
            # is a function of some `n` fraction) but it must produce
            # the same sized array
            sz1, _ = get_arr_size(n1)
            assert sz == sz1, (n, n1, sz, sz1)


class TestModuleDeclaration:
    def test_dependencies(self, tmp_path):
        fpath = util.getpath("tests", "src", "crackfortran", "foo_deps.f90")
        mod = crackfortran.crackfortran([str(fpath)])
        assert len(mod) == 1
        assert mod[0]["vars"]["abar"]["="] == "bar('abar')"


class TestEval(util.F2PyTest):
    def test_eval_scalar(self):
        eval_scalar = crackfortran._eval_scalar

        assert eval_scalar('123', {}) == '123'
        assert eval_scalar('12 + 3', {}) == '15'
        assert eval_scalar('a + b', {"a": 1, "b": 2}) == '3'
        assert eval_scalar('"123"', {}) == "'123'"


class TestFortranReader(util.F2PyTest):
    @pytest.mark.parametrize("encoding",
                             ['ascii', 'utf-8', 'utf-16', 'utf-32'])
    def test_input_encoding(self, tmp_path, encoding):
        # gh-635
        f_path = tmp_path / f"input_with_{encoding}_encoding.f90"
        with f_path.open('w', encoding=encoding) as ff:
            ff.write("""
                     subroutine foo()
                     end subroutine foo
                     """)
        mod = crackfortran.crackfortran([str(f_path)])
        assert mod[0]['name'] == 'foo'


@pytest.mark.slow
class TestUnicodeComment(util.F2PyTest):
    sources = [util.getpath("tests", "src", "crackfortran", "unicode_comment.f90")]

    @pytest.mark.skipif(
        (importlib.util.find_spec("charset_normalizer") is None),
        reason="test requires charset_normalizer which is not installed",
    )
    def test_encoding_comment(self):
        self.module.foo(3)


class TestNameArgsPatternBacktracking:
    @pytest.mark.parametrize(
        ['adversary'],
        [
            ('@)@bind@(@',),
            ('@)@bind                         @(@',),
            ('@)@bind foo bar baz@(@',)
        ]
    )
    def test_nameargspattern_backtracking(self, adversary):
        '''address ReDOS vulnerability:
        https://github.com/numpy/numpy/issues/23338'''
        trials_per_batch = 12
        batches_per_regex = 4
        start_reps, end_reps = 15, 25
        for ii in range(start_reps, end_reps):
            repeated_adversary = adversary * ii
            # test times in small batches.
            # this gives us more chances to catch a bad regex
            # while still catching it before too long if it is bad
            for _ in range(batches_per_regex):
                times = []
                for _ in range(trials_per_batch):
                    t0 = time.perf_counter()
                    mtch = nameargspattern.search(repeated_adversary)
                    times.append(time.perf_counter() - t0)
                # our pattern should be much faster than 0.2s per search
                # it's unlikely that a bad regex will pass even on fast CPUs
                assert np.median(times) < 0.2
            assert not mtch
            # if the adversary is capped with @)@, it becomes acceptable
            # according to the old version of the regex.
            # that should still be true.
            good_version_of_adversary = repeated_adversary + '@)@'
            assert nameargspattern.search(good_version_of_adversary)

class TestFunctionReturn(util.F2PyTest):
    sources = [util.getpath("tests", "src", "crackfortran", "gh23598.f90")]

    @pytest.mark.slow
    def test_function_rettype(self):
        # gh-23598
        assert self.module.intproduct(3, 4) == 12


class TestFortranGroupCounters(util.F2PyTest):
    def test_end_if_comment(self):
        # gh-23533
        fpath = util.getpath("tests", "src", "crackfortran", "gh23533.f")
        try:
            crackfortran.crackfortran([str(fpath)])
        except Exception as exc:
            assert False, f"'crackfortran.crackfortran' raised an exception {exc}"


class TestF77CommonBlockReader:
    def test_gh22648(self, tmp_path):
        fpath = util.getpath("tests", "src", "crackfortran", "gh22648.pyf")
        with contextlib.redirect_stdout(io.StringIO()) as stdout_f2py:
            mod = crackfortran.crackfortran([str(fpath)])
        assert "Mismatch" not in stdout_f2py.getvalue()

class TestParamEval:
    # issue gh-11612, array parameter parsing
    def test_param_eval_nested(self):
        v = '(/3.14, 4./)'
        g_params = {"kind": crackfortran._kind_func,
                "selected_int_kind": crackfortran._selected_int_kind_func,
                "selected_real_kind": crackfortran._selected_real_kind_func}
        params = {'dp': 8, 'intparamarray': {1: 3, 2: 5},
                  'nested': {1: 1, 2: 2, 3: 3}}
        dimspec = '(2)'
        ret = crackfortran.param_eval(v, g_params, params, dimspec=dimspec)
        assert ret == {1: 3.14, 2: 4.0}

    def test_param_eval_nonstandard_range(self):
        v = '(/ 6, 3, 1 /)'
        g_params = {"kind": crackfortran._kind_func,
                "selected_int_kind": crackfortran._selected_int_kind_func,
                "selected_real_kind": crackfortran._selected_real_kind_func}
        params = {}
        dimspec = '(-1:1)'
        ret = crackfortran.param_eval(v, g_params, params, dimspec=dimspec)
        assert ret == {-1: 6, 0: 3, 1: 1}

    def test_param_eval_empty_range(self):
        v = '6'
        g_params = {"kind": crackfortran._kind_func,
                "selected_int_kind": crackfortran._selected_int_kind_func,
                "selected_real_kind": crackfortran._selected_real_kind_func}
        params = {}
        dimspec = ''
        pytest.raises(ValueError, crackfortran.param_eval, v, g_params, params,
                      dimspec=dimspec)

    def test_param_eval_non_array_param(self):
        v = '3.14_dp'
        g_params = {"kind": crackfortran._kind_func,
                "selected_int_kind": crackfortran._selected_int_kind_func,
                "selected_real_kind": crackfortran._selected_real_kind_func}
        params = {}
        ret = crackfortran.param_eval(v, g_params, params, dimspec=None)
        assert ret == '3.14_dp'

    def test_param_eval_too_many_dims(self):
        v = 'reshape((/ (i, i=1, 250) /), (/5, 10, 5/))'
        g_params = {"kind": crackfortran._kind_func,
                "selected_int_kind": crackfortran._selected_int_kind_func,
                "selected_real_kind": crackfortran._selected_real_kind_func}
        params = {}
        dimspec = '(0:4, 3:12, 5)'
        pytest.raises(ValueError, crackfortran.param_eval, v, g_params, params,
                      dimspec=dimspec)

@pytest.mark.slow
class TestLowerF2PYDirective(util.F2PyTest):
    sources = [util.getpath("tests", "src", "crackfortran", "gh27697.f90")]
    options = ['--lower']

    def test_no_lower_fail(self):
        with pytest.raises(ValueError, match='aborting directly') as exc:
            self.module.utils.my_abort('aborting directly')

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\io\sas\test_sas7bdat.py ===
import contextlib
from datetime import datetime
import io
import os
from pathlib import Path

import numpy as np
import pytest

from pandas.compat import IS64
from pandas.errors import EmptyDataError
import pandas.util._test_decorators as td

import pandas as pd
import pandas._testing as tm

from pandas.io.sas.sas7bdat import SAS7BDATReader


@pytest.fixture
def dirpath(datapath):
    return datapath("io", "sas", "data")


@pytest.fixture(params=[(1, range(1, 16)), (2, [16])])
def data_test_ix(request, dirpath):
    i, test_ix = request.param
    fname = os.path.join(dirpath, f"test_sas7bdat_{i}.csv")
    df = pd.read_csv(fname)
    epoch = datetime(1960, 1, 1)
    t1 = pd.to_timedelta(df["Column4"], unit="d")
    df["Column4"] = (epoch + t1).astype("M8[s]")
    t2 = pd.to_timedelta(df["Column12"], unit="d")
    df["Column12"] = (epoch + t2).astype("M8[s]")
    for k in range(df.shape[1]):
        col = df.iloc[:, k]
        if col.dtype == np.int64:
            df.isetitem(k, df.iloc[:, k].astype(np.float64))
    return df, test_ix


# https://github.com/cython/cython/issues/1720
class TestSAS7BDAT:
    @pytest.mark.slow
    def test_from_file(self, dirpath, data_test_ix):
        expected, test_ix = data_test_ix
        for k in test_ix:
            fname = os.path.join(dirpath, f"test{k}.sas7bdat")
            df = pd.read_sas(fname, encoding="utf-8")
            tm.assert_frame_equal(df, expected)

    @pytest.mark.slow
    def test_from_buffer(self, dirpath, data_test_ix):
        expected, test_ix = data_test_ix
        for k in test_ix:
            fname = os.path.join(dirpath, f"test{k}.sas7bdat")
            with open(fname, "rb") as f:
                byts = f.read()
            buf = io.BytesIO(byts)
            with pd.read_sas(
                buf, format="sas7bdat", iterator=True, encoding="utf-8"
            ) as rdr:
                df = rdr.read()
            tm.assert_frame_equal(df, expected)

    @pytest.mark.slow
    def test_from_iterator(self, dirpath, data_test_ix):
        expected, test_ix = data_test_ix
        for k in test_ix:
            fname = os.path.join(dirpath, f"test{k}.sas7bdat")
            with pd.read_sas(fname, iterator=True, encoding="utf-8") as rdr:
                df = rdr.read(2)
                tm.assert_frame_equal(df, expected.iloc[0:2, :])
                df = rdr.read(3)
                tm.assert_frame_equal(df, expected.iloc[2:5, :])

    @pytest.mark.slow
    def test_path_pathlib(self, dirpath, data_test_ix):
        expected, test_ix = data_test_ix
        for k in test_ix:
            fname = Path(os.path.join(dirpath, f"test{k}.sas7bdat"))
            df = pd.read_sas(fname, encoding="utf-8")
            tm.assert_frame_equal(df, expected)

    @td.skip_if_no("py.path")
    @pytest.mark.slow
    def test_path_localpath(self, dirpath, data_test_ix):
        from py.path import local as LocalPath

        expected, test_ix = data_test_ix
        for k in test_ix:
            fname = LocalPath(os.path.join(dirpath, f"test{k}.sas7bdat"))
            df = pd.read_sas(fname, encoding="utf-8")
            tm.assert_frame_equal(df, expected)

    @pytest.mark.slow
    @pytest.mark.parametrize("chunksize", (3, 5, 10, 11))
    @pytest.mark.parametrize("k", range(1, 17))
    def test_iterator_loop(self, dirpath, k, chunksize):
        # github #13654
        fname = os.path.join(dirpath, f"test{k}.sas7bdat")
        with pd.read_sas(fname, chunksize=chunksize, encoding="utf-8") as rdr:
            y = 0
            for x in rdr:
                y += x.shape[0]
        assert y == rdr.row_count

    def test_iterator_read_too_much(self, dirpath):
        # github #14734
        fname = os.path.join(dirpath, "test1.sas7bdat")
        with pd.read_sas(
            fname, format="sas7bdat", iterator=True, encoding="utf-8"
        ) as rdr:
            d1 = rdr.read(rdr.row_count + 20)

        with pd.read_sas(fname, iterator=True, encoding="utf-8") as rdr:
            d2 = rdr.read(rdr.row_count + 20)
        tm.assert_frame_equal(d1, d2)


def test_encoding_options(datapath):
    fname = datapath("io", "sas", "data", "test1.sas7bdat")
    df1 = pd.read_sas(fname)
    df2 = pd.read_sas(fname, encoding="utf-8")
    for col in df1.columns:
        try:
            df1[col] = df1[col].str.decode("utf-8")
        except AttributeError:
            pass
    tm.assert_frame_equal(df1, df2)

    with contextlib.closing(SAS7BDATReader(fname, convert_header_text=False)) as rdr:
        df3 = rdr.read()
    for x, y in zip(df1.columns, df3.columns):
        assert x == y.decode()


def test_encoding_infer(datapath):
    fname = datapath("io", "sas", "data", "test1.sas7bdat")

    with pd.read_sas(fname, encoding="infer", iterator=True) as df1_reader:
        # check: is encoding inferred correctly from file
        assert df1_reader.inferred_encoding == "cp1252"
        df1 = df1_reader.read()

    with pd.read_sas(fname, encoding="cp1252", iterator=True) as df2_reader:
        df2 = df2_reader.read()

    # check: reader reads correct information
    tm.assert_frame_equal(df1, df2)


def test_productsales(datapath):
    fname = datapath("io", "sas", "data", "productsales.sas7bdat")
    df = pd.read_sas(fname, encoding="utf-8")
    fname = datapath("io", "sas", "data", "productsales.csv")
    df0 = pd.read_csv(fname, parse_dates=["MONTH"])
    vn = ["ACTUAL", "PREDICT", "QUARTER", "YEAR"]
    df0[vn] = df0[vn].astype(np.float64)

    df0["MONTH"] = df0["MONTH"].astype("M8[s]")
    tm.assert_frame_equal(df, df0)


def test_12659(datapath):
    fname = datapath("io", "sas", "data", "test_12659.sas7bdat")
    df = pd.read_sas(fname)
    fname = datapath("io", "sas", "data", "test_12659.csv")
    df0 = pd.read_csv(fname)
    df0 = df0.astype(np.float64)
    tm.assert_frame_equal(df, df0)


def test_airline(datapath):
    fname = datapath("io", "sas", "data", "airline.sas7bdat")
    df = pd.read_sas(fname)
    fname = datapath("io", "sas", "data", "airline.csv")
    df0 = pd.read_csv(fname)
    df0 = df0.astype(np.float64)
    tm.assert_frame_equal(df, df0)


def test_date_time(datapath):
    # Support of different SAS date/datetime formats (PR #15871)
    fname = datapath("io", "sas", "data", "datetime.sas7bdat")
    df = pd.read_sas(fname)
    fname = datapath("io", "sas", "data", "datetime.csv")
    df0 = pd.read_csv(
        fname, parse_dates=["Date1", "Date2", "DateTime", "DateTimeHi", "Taiw"]
    )
    # GH 19732: Timestamps imported from sas will incur floating point errors
    # See GH#56014 for discussion of the correct "expected" results
    #  We are really just testing that we are "close". This only seems to be
    #  an issue near the implementation bounds.

    df[df.columns[3]] = df.iloc[:, 3].dt.round("us")
    df0["Date1"] = df0["Date1"].astype("M8[s]")
    df0["Date2"] = df0["Date2"].astype("M8[s]")
    df0["DateTime"] = df0["DateTime"].astype("M8[ms]")
    df0["Taiw"] = df0["Taiw"].astype("M8[s]")

    res = df0["DateTimeHi"].astype("M8[us]").dt.round("ms")
    df0["DateTimeHi"] = res.astype("M8[ms]")

    if not IS64:
        # No good reason for this, just what we get on the CI
        df0.loc[0, "DateTimeHi"] += np.timedelta64(1, "ms")
        df0.loc[[2, 3], "DateTimeHi"] -= np.timedelta64(1, "ms")
    tm.assert_frame_equal(df, df0)


@pytest.mark.parametrize("column", ["WGT", "CYL"])
def test_compact_numerical_values(datapath, column):
    # Regression test for #21616
    fname = datapath("io", "sas", "data", "cars.sas7bdat")
    df = pd.read_sas(fname, encoding="latin-1")
    # The two columns CYL and WGT in cars.sas7bdat have column
    # width < 8 and only contain integral values.
    # Test that pandas doesn't corrupt the numbers by adding
    # decimals.
    result = df[column]
    expected = df[column].round()
    tm.assert_series_equal(result, expected, check_exact=True)


def test_many_columns(datapath):
    # Test for looking for column information in more places (PR #22628)
    fname = datapath("io", "sas", "data", "many_columns.sas7bdat")

    df = pd.read_sas(fname, encoding="latin-1")

    fname = datapath("io", "sas", "data", "many_columns.csv")
    df0 = pd.read_csv(fname, encoding="latin-1")
    tm.assert_frame_equal(df, df0)


def test_inconsistent_number_of_rows(datapath):
    # Regression test for issue #16615. (PR #22628)
    fname = datapath("io", "sas", "data", "load_log.sas7bdat")
    df = pd.read_sas(fname, encoding="latin-1")
    assert len(df) == 2097


def test_zero_variables(datapath):
    # Check if the SAS file has zero variables (PR #18184)
    fname = datapath("io", "sas", "data", "zero_variables.sas7bdat")
    with pytest.raises(EmptyDataError, match="No columns to parse from file"):
        pd.read_sas(fname)


@pytest.mark.parametrize("encoding", [None, "utf8"])
def test_zero_rows(datapath, encoding):
    # GH 18198
    fname = datapath("io", "sas", "data", "zero_rows.sas7bdat")
    result = pd.read_sas(fname, encoding=encoding)
    str_value = b"a" if encoding is None else "a"
    expected = pd.DataFrame([{"char_field": str_value, "num_field": 1.0}]).iloc[:0]
    tm.assert_frame_equal(result, expected)


def test_corrupt_read(datapath):
    # We don't really care about the exact failure, the important thing is
    # that the resource should be cleaned up afterwards (BUG #35566)
    fname = datapath("io", "sas", "data", "corrupt.sas7bdat")
    msg = "'SAS7BDATReader' object has no attribute 'row_count'"
    with pytest.raises(AttributeError, match=msg):
        pd.read_sas(fname)


def test_max_sas_date(datapath):
    # GH 20927
    # NB. max datetime in SAS dataset is 31DEC9999:23:59:59.999
    #    but this is read as 29DEC9999:23:59:59.998993 by a buggy
    #    sas7bdat module
    # See also GH#56014 for discussion of the correct "expected" results.
    fname = datapath("io", "sas", "data", "max_sas_date.sas7bdat")
    df = pd.read_sas(fname, encoding="iso-8859-1")

    expected = pd.DataFrame(
        {
            "text": ["max", "normal"],
            "dt_as_float": [253717747199.999, 1880323199.999],
            "dt_as_dt": np.array(
                [
                    datetime(9999, 12, 29, 23, 59, 59, 999000),
                    datetime(2019, 8, 1, 23, 59, 59, 999000),
                ],
                dtype="M8[ms]",
            ),
            "date_as_float": [2936547.0, 21762.0],
            "date_as_date": np.array(
                [
                    datetime(9999, 12, 29),
                    datetime(2019, 8, 1),
                ],
                dtype="M8[s]",
            ),
        },
        columns=["text", "dt_as_float", "dt_as_dt", "date_as_float", "date_as_date"],
    )

    if not IS64:
        # No good reason for this, just what we get on the CI
        expected.loc[:, "dt_as_dt"] -= np.timedelta64(1, "ms")

    tm.assert_frame_equal(df, expected)


def test_max_sas_date_iterator(datapath):
    # GH 20927
    # when called as an iterator, only those chunks with a date > pd.Timestamp.max
    # are returned as datetime.datetime, if this happens that whole chunk is returned
    # as datetime.datetime
    col_order = ["text", "dt_as_float", "dt_as_dt", "date_as_float", "date_as_date"]
    fname = datapath("io", "sas", "data", "max_sas_date.sas7bdat")
    results = []
    for df in pd.read_sas(fname, encoding="iso-8859-1", chunksize=1):
        # GH 19732: Timestamps imported from sas will incur floating point errors
        df.reset_index(inplace=True, drop=True)
        results.append(df)
    expected = [
        pd.DataFrame(
            {
                "text": ["max"],
                "dt_as_float": [253717747199.999],
                "dt_as_dt": np.array(
                    [datetime(9999, 12, 29, 23, 59, 59, 999000)], dtype="M8[ms]"
                ),
                "date_as_float": [2936547.0],
                "date_as_date": np.array([datetime(9999, 12, 29)], dtype="M8[s]"),
            },
            columns=col_order,
        ),
        pd.DataFrame(
            {
                "text": ["normal"],
                "dt_as_float": [1880323199.999],
                "dt_as_dt": np.array(["2019-08-01 23:59:59.999"], dtype="M8[ms]"),
                "date_as_float": [21762.0],
                "date_as_date": np.array(["2019-08-01"], dtype="M8[s]"),
            },
            columns=col_order,
        ),
    ]
    if not IS64:
        # No good reason for this, just what we get on the CI
        expected[0].loc[0, "dt_as_dt"] -= np.timedelta64(1, "ms")
        expected[1].loc[0, "dt_as_dt"] -= np.timedelta64(1, "ms")

    tm.assert_frame_equal(results[0], expected[0])
    tm.assert_frame_equal(results[1], expected[1])


def test_null_date(datapath):
    fname = datapath("io", "sas", "data", "dates_null.sas7bdat")
    df = pd.read_sas(fname, encoding="utf-8")

    expected = pd.DataFrame(
        {
            "datecol": np.array(
                [
                    datetime(9999, 12, 29),
                    np.datetime64("NaT"),
                ],
                dtype="M8[s]",
            ),
            "datetimecol": np.array(
                [
                    datetime(9999, 12, 29, 23, 59, 59, 999000),
                    np.datetime64("NaT"),
                ],
                dtype="M8[ms]",
            ),
        },
    )
    if not IS64:
        # No good reason for this, just what we get on the CI
        expected.loc[0, "datetimecol"] -= np.timedelta64(1, "ms")
    tm.assert_frame_equal(df, expected)


def test_meta2_page(datapath):
    # GH 35545
    fname = datapath("io", "sas", "data", "test_meta2_page.sas7bdat")
    df = pd.read_sas(fname)
    assert len(df) == 1000


@pytest.mark.parametrize(
    "test_file, override_offset, override_value, expected_msg",
    [
        ("test2.sas7bdat", 0x10000 + 55229, 0x80 | 0x0F, "Out of bounds"),
        ("test2.sas7bdat", 0x10000 + 55229, 0x10, "unknown control byte"),
        ("test3.sas7bdat", 118170, 184, "Out of bounds"),
    ],
)
def test_rle_rdc_exceptions(
    datapath, test_file, override_offset, override_value, expected_msg
):
    """Errors in RLE/RDC decompression should propagate."""
    with open(datapath("io", "sas", "data", test_file), "rb") as fd:
        data = bytearray(fd.read())
    data[override_offset] = override_value
    with pytest.raises(Exception, match=expected_msg):
        pd.read_sas(io.BytesIO(data), format="sas7bdat")


def test_0x40_control_byte(datapath):
    # GH 31243
    fname = datapath("io", "sas", "data", "0x40controlbyte.sas7bdat")
    df = pd.read_sas(fname, encoding="ascii")
    fname = datapath("io", "sas", "data", "0x40controlbyte.csv")
    df0 = pd.read_csv(fname, dtype="str")
    tm.assert_frame_equal(df, df0)


def test_0x00_control_byte(datapath):
    # GH 47099
    fname = datapath("io", "sas", "data", "0x00controlbyte.sas7bdat.bz2")
    df = next(pd.read_sas(fname, chunksize=11_000))
    assert df.shape == (11_000, 20)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\core\tests\test_symbol.py ===
import threading

from sympy.core.function import Function, UndefinedFunction
from sympy.core.numbers import (I, Rational, pi)
from sympy.core.relational import (GreaterThan, LessThan, StrictGreaterThan, StrictLessThan)
from sympy.core.symbol import (Dummy, Symbol, Wild, symbols)
from sympy.core.sympify import sympify  # can't import as S yet
from sympy.core.symbol import uniquely_named_symbol, _symbol, Str

from sympy.testing.pytest import raises, skip_under_pyodide
from sympy.core.symbol import disambiguate


def test_Str():
    a1 = Str('a')
    a2 = Str('a')
    b = Str('b')
    assert a1 == a2 != b
    raises(TypeError, lambda: Str())


def test_Symbol():
    a = Symbol("a")
    x1 = Symbol("x")
    x2 = Symbol("x")
    xdummy1 = Dummy("x")
    xdummy2 = Dummy("x")

    assert a != x1
    assert a != x2
    assert x1 == x2
    assert x1 != xdummy1
    assert xdummy1 != xdummy2

    assert Symbol("x") == Symbol("x")
    assert Dummy("x") != Dummy("x")
    d = symbols('d', cls=Dummy)
    assert isinstance(d, Dummy)
    c, d = symbols('c,d', cls=Dummy)
    assert isinstance(c, Dummy)
    assert isinstance(d, Dummy)
    raises(TypeError, lambda: Symbol())


def test_Dummy():
    assert Dummy() != Dummy()


def test_Dummy_force_dummy_index():
    raises(AssertionError, lambda: Dummy(dummy_index=1))
    assert Dummy('d', dummy_index=2) == Dummy('d', dummy_index=2)
    assert Dummy('d1', dummy_index=2) != Dummy('d2', dummy_index=2)
    d1 = Dummy('d', dummy_index=3)
    d2 = Dummy('d')
    # might fail if d1 were created with dummy_index >= 10**6
    assert d1 != d2
    d3 = Dummy('d', dummy_index=3)
    assert d1 == d3
    assert Dummy()._count == Dummy('d', dummy_index=3)._count


def test_lt_gt():
    S = sympify
    x, y = Symbol('x'), Symbol('y')

    assert (x >= y) == GreaterThan(x, y)
    assert (x >= 0) == GreaterThan(x, 0)
    assert (x <= y) == LessThan(x, y)
    assert (x <= 0) == LessThan(x, 0)

    assert (0 <= x) == GreaterThan(x, 0)
    assert (0 >= x) == LessThan(x, 0)
    assert (S(0) >= x) == GreaterThan(0, x)
    assert (S(0) <= x) == LessThan(0, x)

    assert (x > y) == StrictGreaterThan(x, y)
    assert (x > 0) == StrictGreaterThan(x, 0)
    assert (x < y) == StrictLessThan(x, y)
    assert (x < 0) == StrictLessThan(x, 0)

    assert (0 < x) == StrictGreaterThan(x, 0)
    assert (0 > x) == StrictLessThan(x, 0)
    assert (S(0) > x) == StrictGreaterThan(0, x)
    assert (S(0) < x) == StrictLessThan(0, x)

    e = x**2 + 4*x + 1
    assert (e >= 0) == GreaterThan(e, 0)
    assert (0 <= e) == GreaterThan(e, 0)
    assert (e > 0) == StrictGreaterThan(e, 0)
    assert (0 < e) == StrictGreaterThan(e, 0)

    assert (e <= 0) == LessThan(e, 0)
    assert (0 >= e) == LessThan(e, 0)
    assert (e < 0) == StrictLessThan(e, 0)
    assert (0 > e) == StrictLessThan(e, 0)

    assert (S(0) >= e) == GreaterThan(0, e)
    assert (S(0) <= e) == LessThan(0, e)
    assert (S(0) < e) == StrictLessThan(0, e)
    assert (S(0) > e) == StrictGreaterThan(0, e)


def test_no_len():
    # there should be no len for numbers
    x = Symbol('x')
    raises(TypeError, lambda: len(x))


def test_ineq_unequal():
    S = sympify
    x, y, z = symbols('x,y,z')

    e = (
        S(-1) >= x, S(-1) >= y, S(-1) >= z,
        S(-1) > x, S(-1) > y, S(-1) > z,
        S(-1) <= x, S(-1) <= y, S(-1) <= z,
        S(-1) < x, S(-1) < y, S(-1) < z,
        S(0) >= x, S(0) >= y, S(0) >= z,
        S(0) > x, S(0) > y, S(0) > z,
        S(0) <= x, S(0) <= y, S(0) <= z,
        S(0) < x, S(0) < y, S(0) < z,
        S('3/7') >= x, S('3/7') >= y, S('3/7') >= z,
        S('3/7') > x, S('3/7') > y, S('3/7') > z,
        S('3/7') <= x, S('3/7') <= y, S('3/7') <= z,
        S('3/7') < x, S('3/7') < y, S('3/7') < z,
        S(1.5) >= x, S(1.5) >= y, S(1.5) >= z,
        S(1.5) > x, S(1.5) > y, S(1.5) > z,
        S(1.5) <= x, S(1.5) <= y, S(1.5) <= z,
        S(1.5) < x, S(1.5) < y, S(1.5) < z,
        S(2) >= x, S(2) >= y, S(2) >= z,
        S(2) > x, S(2) > y, S(2) > z,
        S(2) <= x, S(2) <= y, S(2) <= z,
        S(2) < x, S(2) < y, S(2) < z,
        x >= -1, y >= -1, z >= -1,
        x > -1, y > -1, z > -1,
        x <= -1, y <= -1, z <= -1,
        x < -1, y < -1, z < -1,
        x >= 0, y >= 0, z >= 0,
        x > 0, y > 0, z > 0,
        x <= 0, y <= 0, z <= 0,
        x < 0, y < 0, z < 0,
        x >= 1.5, y >= 1.5, z >= 1.5,
        x > 1.5, y > 1.5, z > 1.5,
        x <= 1.5, y <= 1.5, z <= 1.5,
        x < 1.5, y < 1.5, z < 1.5,
        x >= 2, y >= 2, z >= 2,
        x > 2, y > 2, z > 2,
        x <= 2, y <= 2, z <= 2,
        x < 2, y < 2, z < 2,

        x >= y, x >= z, y >= x, y >= z, z >= x, z >= y,
        x > y, x > z, y > x, y > z, z > x, z > y,
        x <= y, x <= z, y <= x, y <= z, z <= x, z <= y,
        x < y, x < z, y < x, y < z, z < x, z < y,

        x - pi >= y + z, y - pi >= x + z, z - pi >= x + y,
        x - pi > y + z, y - pi > x + z, z - pi > x + y,
        x - pi <= y + z, y - pi <= x + z, z - pi <= x + y,
        x - pi < y + z, y - pi < x + z, z - pi < x + y,
        True, False
    )

    left_e = e[:-1]
    for i, e1 in enumerate( left_e ):
        for e2 in e[i + 1:]:
            assert e1 != e2


def test_Wild_properties():
    S = sympify
    # these tests only include Atoms
    x = Symbol("x")
    y = Symbol("y")
    p = Symbol("p", positive=True)
    k = Symbol("k", integer=True)
    n = Symbol("n", integer=True, positive=True)

    given_patterns = [ x, y, p, k, -k, n, -n, S(-3), S(3),
                       pi, Rational(3, 2), I ]

    integerp = lambda k: k.is_integer
    positivep = lambda k: k.is_positive
    symbolp = lambda k: k.is_Symbol
    realp = lambda k: k.is_extended_real

    S = Wild("S", properties=[symbolp])
    R = Wild("R", properties=[realp])
    Y = Wild("Y", exclude=[x, p, k, n])
    P = Wild("P", properties=[positivep])
    K = Wild("K", properties=[integerp])
    N = Wild("N", properties=[positivep, integerp])

    given_wildcards = [ S, R, Y, P, K, N ]

    goodmatch = {
        S: (x, y, p, k, n),
        R: (p, k, -k, n, -n, -3, 3, pi, Rational(3, 2)),
        Y: (y, -3, 3, pi, Rational(3, 2), I ),
        P: (p, n, 3, pi, Rational(3, 2)),
        K: (k, -k, n, -n, -3, 3),
        N: (n, 3)}

    for A in given_wildcards:
        for pat in given_patterns:
            d = pat.match(A)
            if pat in goodmatch[A]:
                assert d[A] in goodmatch[A]
            else:
                assert d is None


def test_symbols():
    x = Symbol('x')
    y = Symbol('y')
    z = Symbol('z')

    assert symbols('x') == x
    assert symbols('x ') == x
    assert symbols(' x ') == x
    assert symbols('x,') == (x,)
    assert symbols('x, ') == (x,)
    assert symbols('x ,') == (x,)

    assert symbols('x , y') == (x, y)

    assert symbols('x,y,z') == (x, y, z)
    assert symbols('x y z') == (x, y, z)

    assert symbols('x,y,z,') == (x, y, z)
    assert symbols('x y z ') == (x, y, z)

    xyz = Symbol('xyz')
    abc = Symbol('abc')

    assert symbols('xyz') == xyz
    assert symbols('xyz,') == (xyz,)
    assert symbols('xyz,abc') == (xyz, abc)

    assert symbols(('xyz',)) == (xyz,)
    assert symbols(('xyz,',)) == ((xyz,),)
    assert symbols(('x,y,z,',)) == ((x, y, z),)
    assert symbols(('xyz', 'abc')) == (xyz, abc)
    assert symbols(('xyz,abc',)) == ((xyz, abc),)
    assert symbols(('xyz,abc', 'x,y,z')) == ((xyz, abc), (x, y, z))

    assert symbols(('x', 'y', 'z')) == (x, y, z)
    assert symbols(['x', 'y', 'z']) == [x, y, z]
    assert symbols({'x', 'y', 'z'}) == {x, y, z}

    raises(ValueError, lambda: symbols(''))
    raises(ValueError, lambda: symbols(','))
    raises(ValueError, lambda: symbols('x,,y,,z'))
    raises(ValueError, lambda: symbols(('x', '', 'y', '', 'z')))

    a, b = symbols('x,y', real=True)
    assert a.is_real and b.is_real

    x0 = Symbol('x0')
    x1 = Symbol('x1')
    x2 = Symbol('x2')

    y0 = Symbol('y0')
    y1 = Symbol('y1')

    assert symbols('x0:0') == ()
    assert symbols('x0:1') == (x0,)
    assert symbols('x0:2') == (x0, x1)
    assert symbols('x0:3') == (x0, x1, x2)

    assert symbols('x:0') == ()
    assert symbols('x:1') == (x0,)
    assert symbols('x:2') == (x0, x1)
    assert symbols('x:3') == (x0, x1, x2)

    assert symbols('x1:1') == ()
    assert symbols('x1:2') == (x1,)
    assert symbols('x1:3') == (x1, x2)

    assert symbols('x1:3,x,y,z') == (x1, x2, x, y, z)

    assert symbols('x:3,y:2') == (x0, x1, x2, y0, y1)
    assert symbols(('x:3', 'y:2')) == ((x0, x1, x2), (y0, y1))

    a = Symbol('a')
    b = Symbol('b')
    c = Symbol('c')
    d = Symbol('d')

    assert symbols('x:z') == (x, y, z)
    assert symbols('a:d,x:z') == (a, b, c, d, x, y, z)
    assert symbols(('a:d', 'x:z')) == ((a, b, c, d), (x, y, z))

    aa = Symbol('aa')
    ab = Symbol('ab')
    ac = Symbol('ac')
    ad = Symbol('ad')

    assert symbols('aa:d') == (aa, ab, ac, ad)
    assert symbols('aa:d,x:z') == (aa, ab, ac, ad, x, y, z)
    assert symbols(('aa:d','x:z')) == ((aa, ab, ac, ad), (x, y, z))

    assert type(symbols(('q:2', 'u:2'), cls=Function)[0][0]) == UndefinedFunction  # issue 23532

    # issue 6675
    def sym(s):
        return str(symbols(s))
    assert sym('a0:4') == '(a0, a1, a2, a3)'
    assert sym('a2:4,b1:3') == '(a2, a3, b1, b2)'
    assert sym('a1(2:4)') == '(a12, a13)'
    assert sym('a0:2.0:2') == '(a0.0, a0.1, a1.0, a1.1)'
    assert sym('aa:cz') == '(aaz, abz, acz)'
    assert sym('aa:c0:2') == '(aa0, aa1, ab0, ab1, ac0, ac1)'
    assert sym('aa:ba:b') == '(aaa, aab, aba, abb)'
    assert sym('a:3b') == '(a0b, a1b, a2b)'
    assert sym('a-1:3b') == '(a-1b, a-2b)'
    assert sym(r'a:2\,:2' + chr(0)) == '(a0,0%s, a0,1%s, a1,0%s, a1,1%s)' % (
        (chr(0),)*4)
    assert sym('x(:a:3)') == '(x(a0), x(a1), x(a2))'
    assert sym('x(:c):1') == '(xa0, xb0, xc0)'
    assert sym('x((:a)):3') == '(x(a)0, x(a)1, x(a)2)'
    assert sym('x(:a:3') == '(x(a0, x(a1, x(a2)'
    assert sym(':2') == '(0, 1)'
    assert sym(':b') == '(a, b)'
    assert sym(':b:2') == '(a0, a1, b0, b1)'
    assert sym(':2:2') == '(00, 01, 10, 11)'
    assert sym(':b:b') == '(aa, ab, ba, bb)'

    raises(ValueError, lambda: symbols(':'))
    raises(ValueError, lambda: symbols('a:'))
    raises(ValueError, lambda: symbols('::'))
    raises(ValueError, lambda: symbols('a::'))
    raises(ValueError, lambda: symbols(':a:'))
    raises(ValueError, lambda: symbols('::a'))


def test_symbols_become_functions_issue_3539():
    from sympy.abc import alpha, phi, beta, t
    raises(TypeError, lambda: beta(2))
    raises(TypeError, lambda: beta(2.5))
    raises(TypeError, lambda: phi(2.5))
    raises(TypeError, lambda: alpha(2.5))
    raises(TypeError, lambda: phi(t))


def test_unicode():
    xu = Symbol('x')
    x = Symbol('x')
    assert x == xu

    raises(TypeError, lambda: Symbol(1))


def test_uniquely_named_symbol_and_Symbol():
    F = uniquely_named_symbol
    x = Symbol('x')
    assert F(x) == x
    assert F('x') == x
    assert str(F('x', x)) == 'x0'
    assert str(F('x', (x + 1, 1/x))) == 'x0'
    _x = Symbol('x', real=True)
    assert F(('x', _x)) == _x
    assert F((x, _x)) == _x
    assert F('x', real=True).is_real
    y = Symbol('y')
    assert F(('x', y), real=True).is_real
    r = Symbol('x', real=True)
    assert F(('x', r)).is_real
    assert F(('x', r), real=False).is_real
    assert F('x1', Symbol('x1'),
        compare=lambda i: str(i).rstrip('1')).name == 'x0'
    assert F('x1', Symbol('x1'),
        modify=lambda i: i + '_').name == 'x1_'
    assert _symbol(x, _x) == x


def test_disambiguate():
    x, y, y_1, _x, x_1, x_2 = symbols('x y y_1 _x x_1 x_2')
    t1 = Dummy('y'), _x, Dummy('x'), Dummy('x')
    t2 = Dummy('x'), Dummy('x')
    t3 = Dummy('x'), Dummy('y')
    t4 = x, Dummy('x')
    t5 = Symbol('x', integer=True), x, Symbol('x_1')

    assert disambiguate(*t1) == (y, x_2, x, x_1)
    assert disambiguate(*t2) == (x, x_1)
    assert disambiguate(*t3) == (x, y)
    assert disambiguate(*t4) == (x_1, x)
    assert disambiguate(*t5) == (t5[0], x_2, x_1)
    assert disambiguate(*t5)[0] != x  # assumptions are retained

    t6 = _x, Dummy('x')/y
    t7 = y*Dummy('y'), y

    assert disambiguate(*t6) == (x_1, x/y)
    assert disambiguate(*t7) == (y*y_1, y_1)
    assert disambiguate(Dummy('x_1'), Dummy('x_1')
        ) == (x_1, Symbol('x_1_1'))


@skip_under_pyodide("Cannot create threads under pyodide.")
def test_issue_gh_16734():
    # https://github.com/sympy/sympy/issues/16734

    syms = list(symbols('x, y'))

    def thread1():
        for n in range(1000):
            syms[0], syms[1] = symbols(f'x{n}, y{n}')
            syms[0].is_positive # Check an assumption in this thread.
        syms[0] = None

    def thread2():
        while syms[0] is not None:
            # Compare the symbol in this thread.
            result = (syms[0] == syms[1])  # noqa

    # Previously this would be very likely to raise an exception:
    thread = threading.Thread(target=thread1)
    thread.start()
    thread2()
    thread.join()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\series\tests\test_series.py ===
from sympy.core.evalf import N
from sympy.core.function import (Derivative, Function, PoleError, Subs)
from sympy.core.numbers import (E, Float, Rational, oo, pi, I)
from sympy.core.singleton import S
from sympy.core.symbol import (Symbol, symbols)
from sympy.functions.elementary.exponential import (LambertW, exp, log)
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.trigonometric import (atan, cos, sin)
from sympy.functions.special.gamma_functions import gamma
from sympy.integrals.integrals import Integral, integrate
from sympy.series.order import O
from sympy.series.series import series
from sympy.abc import x, y, n, k
from sympy.testing.pytest import raises
from sympy.core import EulerGamma


def test_sin():
    e1 = sin(x).series(x, 0)
    e2 = series(sin(x), x, 0)
    assert e1 == e2


def test_cos():
    e1 = cos(x).series(x, 0)
    e2 = series(cos(x), x, 0)
    assert e1 == e2


def test_exp():
    e1 = exp(x).series(x, 0)
    e2 = series(exp(x), x, 0)
    assert e1 == e2


def test_exp2():
    e1 = exp(cos(x)).series(x, 0)
    e2 = series(exp(cos(x)), x, 0)
    assert e1 == e2


def test_issue_5223():
    assert series(1, x) == 1
    assert next(S.Zero.lseries(x)) == 0
    assert cos(x).series() == cos(x).series(x)
    raises(ValueError, lambda: cos(x + y).series())
    raises(ValueError, lambda: x.series(dir=""))

    assert (cos(x).series(x, 1) -
            cos(x + 1).series(x).subs(x, x - 1)).removeO() == 0
    e = cos(x).series(x, 1, n=None)
    assert [next(e) for i in range(2)] == [cos(1), -((x - 1)*sin(1))]
    e = cos(x).series(x, 1, n=None, dir='-')
    assert [next(e) for i in range(2)] == [cos(1), (1 - x)*sin(1)]
    # the following test is exact so no need for x -> x - 1 replacement
    assert abs(x).series(x, 1, dir='-') == x
    assert exp(x).series(x, 1, dir='-', n=3).removeO() == \
        E - E*(-x + 1) + E*(-x + 1)**2/2

    D = Derivative
    assert D(x**2 + x**3*y**2, x, 2, y, 1).series(x).doit() == 12*x*y
    assert next(D(cos(x), x).lseries()) == D(1, x)
    assert D(
        exp(x), x).series(n=3) == D(1, x) + D(x, x) + D(x**2/2, x) + D(x**3/6, x) + O(x**3)

    assert Integral(x, (x, 1, 3), (y, 1, x)).series(x) == -4 + 4*x

    assert (1 + x + O(x**2)).getn() == 2
    assert (1 + x).getn() is None

    raises(PoleError, lambda: ((1/sin(x))**oo).series())
    logx = Symbol('logx')
    assert ((sin(x))**y).nseries(x, n=1, logx=logx) == \
        exp(y*logx) + O(x*exp(y*logx), x)

    assert sin(1/x).series(x, oo, n=5) == 1/x - 1/(6*x**3) + O(x**(-5), (x, oo))
    assert abs(x).series(x, oo, n=5, dir='+') == x
    assert abs(x).series(x, -oo, n=5, dir='-') == -x
    assert abs(-x).series(x, oo, n=5, dir='+') == x
    assert abs(-x).series(x, -oo, n=5, dir='-') == -x

    assert exp(x*log(x)).series(n=3) == \
        1 + x*log(x) + x**2*log(x)**2/2 + O(x**3*log(x)**3)
    # XXX is this right? If not, fix "ngot > n" handling in expr.
    p = Symbol('p', positive=True)
    assert exp(sqrt(p)**3*log(p)).series(n=3) == \
        1 + p**S('3/2')*log(p) + O(p**3*log(p)**3)

    assert exp(sin(x)*log(x)).series(n=2) == 1 + x*log(x) + O(x**2*log(x)**2)


def test_issue_6350():
    expr = integrate(exp(k*(y**3 - 3*y)), (y, 0, oo), conds='none')
    assert expr.series(k, 0, 3) == -(-1)**(S(2)/3)*sqrt(3)*gamma(S(1)/3)**2*gamma(S(2)/3)/(6*pi*k**(S(1)/3)) - \
        sqrt(3)*k*gamma(-S(2)/3)*gamma(-S(1)/3)/(6*pi) - \
        (-1)**(S(1)/3)*sqrt(3)*k**(S(1)/3)*gamma(-S(1)/3)*gamma(S(1)/3)*gamma(S(2)/3)/(6*pi) - \
        (-1)**(S(2)/3)*sqrt(3)*k**(S(5)/3)*gamma(S(1)/3)**2*gamma(S(2)/3)/(4*pi) - \
        (-1)**(S(1)/3)*sqrt(3)*k**(S(7)/3)*gamma(-S(1)/3)*gamma(S(1)/3)*gamma(S(2)/3)/(8*pi) + O(k**3)


def test_issue_11313():
    assert Integral(cos(x), x).series(x) == sin(x).series(x)
    assert Derivative(sin(x), x).series(x, n=3).doit() == cos(x).series(x, n=3)

    assert Derivative(x**3, x).as_leading_term(x) == 3*x**2
    assert Derivative(x**3, y).as_leading_term(x) == 0
    assert Derivative(sin(x), x).as_leading_term(x) == 1
    assert Derivative(cos(x), x).as_leading_term(x) == -x

    # This result is equivalent to zero, zero is not return because
    # `Expr.series` doesn't currently detect an `x` in its `free_symbol`s.
    assert Derivative(1, x).as_leading_term(x) == Derivative(1, x)

    assert Derivative(exp(x), x).series(x).doit() == exp(x).series(x)
    assert 1 + Integral(exp(x), x).series(x) == exp(x).series(x)

    assert Derivative(log(x), x).series(x).doit() == (1/x).series(x)
    assert Integral(log(x), x).series(x) == Integral(log(x), x).doit().series(x).removeO()


def test_series_of_Subs():
    from sympy.abc import z

    subs1 = Subs(sin(x), x, y)
    subs2 = Subs(sin(x) * cos(z), x, y)
    subs3 = Subs(sin(x * z), (x, z), (y, x))

    assert subs1.series(x) == subs1
    subs1_series = (Subs(x, x, y) + Subs(-x**3/6, x, y) +
        Subs(x**5/120, x, y) + O(y**6))
    assert subs1.series() == subs1_series
    assert subs1.series(y) == subs1_series
    assert subs1.series(z) == subs1
    assert subs2.series(z) == (Subs(z**4*sin(x)/24, x, y) +
        Subs(-z**2*sin(x)/2, x, y) + Subs(sin(x), x, y) + O(z**6))
    assert subs3.series(x).doit() == subs3.doit().series(x)
    assert subs3.series(z).doit() == sin(x*y)

    raises(ValueError, lambda: Subs(x + 2*y, y, z).series())
    assert Subs(x + y, y, z).series(x).doit() == x + z


def test_issue_3978():
    f = Function('f')
    assert f(x).series(x, 0, 3, dir='-') == \
            f(0) + x*Subs(Derivative(f(x), x), x, 0) + \
            x**2*Subs(Derivative(f(x), x, x), x, 0)/2 + O(x**3)
    assert f(x).series(x, 0, 3) == \
            f(0) + x*Subs(Derivative(f(x), x), x, 0) + \
            x**2*Subs(Derivative(f(x), x, x), x, 0)/2 + O(x**3)
    assert f(x**2).series(x, 0, 3) == \
            f(0) + x**2*Subs(Derivative(f(x), x), x, 0) + O(x**3)
    assert f(x**2+1).series(x, 0, 3) == \
            f(1) + x**2*Subs(Derivative(f(x), x), x, 1) + O(x**3)

    class TestF(Function):
        pass

    assert TestF(x).series(x, 0, 3) ==  TestF(0) + \
            x*Subs(Derivative(TestF(x), x), x, 0) + \
            x**2*Subs(Derivative(TestF(x), x, x), x, 0)/2 + O(x**3)

from sympy.series.acceleration import richardson, shanks
from sympy.concrete.summations import Sum
from sympy.core.numbers import Integer


def test_acceleration():
    e = (1 + 1/n)**n
    assert round(richardson(e, n, 10, 20).evalf(), 10) == round(E.evalf(), 10)

    A = Sum(Integer(-1)**(k + 1) / k, (k, 1, n))
    assert round(shanks(A, n, 25).evalf(), 4) == round(log(2).evalf(), 4)
    assert round(shanks(A, n, 25, 5).evalf(), 10) == round(log(2).evalf(), 10)


def test_issue_5852():
    assert series(1/cos(x/log(x)), x, 0) == 1 + x**2/(2*log(x)**2) + \
        5*x**4/(24*log(x)**4) + O(x**6)


def test_issue_4583():
    assert cos(1 + x + x**2).series(x, 0, 5) == cos(1) - x*sin(1) + \
        x**2*(-sin(1) - cos(1)/2) + x**3*(-cos(1) + sin(1)/6) + \
        x**4*(-11*cos(1)/24 + sin(1)/2) + O(x**5)


def test_issue_6318():
    eq = (1/x)**Rational(2, 3)
    assert (eq + 1).as_leading_term(x) == eq


def test_x_is_base_detection():
    eq = (x**2)**Rational(2, 3)
    assert eq.series() == x**Rational(4, 3)


def test_issue_7203():
    assert series(cos(x), x, pi, 3) == \
        -1 + (x - pi)**2/2 + O((x - pi)**3, (x, pi))


def test_exp_product_positive_factors():
    a, b = symbols('a, b', positive=True)
    x = a * b
    assert series(exp(x), x, n=8) == 1 + a*b + a**2*b**2/2 + \
        a**3*b**3/6 + a**4*b**4/24 + a**5*b**5/120 + a**6*b**6/720 + \
        a**7*b**7/5040 + O(a**8*b**8, a, b)


def test_issue_8805():
    assert series(1, n=8) == 1


def test_issue_9173():
    p0,p1,p2,p3,b0,b1,b2=symbols('p0 p1 p2 p3 b0 b1 b2')
    Q=(p0+(p1+(p2+p3/y)/y)/y)/(1+((p3/(b0*y)+(b0*p2-b1*p3)/b0**2)/y+\
       (b0**2*p1-b0*b1*p2-p3*(b0*b2-b1**2))/b0**3)/y)

    series = Q.series(y,n=3)

    assert series == y*(b0*p2/p3+b0*(-p2/p3+b1/b0))+y**2*(b0*p1/p3+b0*p2*\
            (-p2/p3+b1/b0)/p3+b0*(-p1/p3+(p2/p3-b1/b0)**2+b1*p2/(b0*p3)+\
            b2/b0-b1**2/b0**2))+b0+O(y**3)
    assert series.simplify() == b2*y**2 + b1*y + b0 + O(y**3)


def test_issue_9549():
    y = (x**2 + x + 1) / (x**3 + x**2)
    assert series(y, x, oo) == x**(-5) - 1/x**4 + x**(-3) + 1/x + O(x**(-6), (x, oo))


def test_issue_10761():
    assert series(1/(x**-2 + x**-3), x, 0) == x**3 - x**4 + x**5 + O(x**6)


def test_issue_12578():
    y = (1 - 1/(x/2 - 1/(2*x))**4)**(S(1)/8)
    assert y.series(x, 0, n=17) == 1 - 2*x**4 - 8*x**6 - 34*x**8 - 152*x**10 - 714*x**12 - \
        3472*x**14 - 17318*x**16 + O(x**17)


def test_issue_12791():
    beta = symbols('beta', positive=True)
    theta, varphi = symbols('theta varphi', real=True)

    expr = (-beta**2*varphi*sin(theta) + beta**2*cos(theta) + \
        beta*varphi*sin(theta) - beta*cos(theta) - beta + 1)/(beta*cos(theta) - 1)**2

    sol = (0.5/(0.5*cos(theta) - 1.0)**2 - 0.25*cos(theta)/(0.5*cos(theta) - 1.0)**2
        + (beta - 0.5)*(-0.25*varphi*sin(2*theta) - 1.5*cos(theta)
        + 0.25*cos(2*theta) + 1.25)/((0.5*cos(theta) - 1.0)**2*(0.5*cos(theta) - 1.0))
        + 0.25*varphi*sin(theta)/(0.5*cos(theta) - 1.0)**2
        + O((beta - S.Half)**2, (beta, S.Half)))

    assert expr.series(beta, 0.5, 2).trigsimp() == sol


def test_issue_14384():
    x, a = symbols('x a')
    assert series(x**a, x) == x**a
    assert series(x**(-2*a), x) == x**(-2*a)
    assert series(exp(a*log(x)), x) == exp(a*log(x))
    raises(PoleError, lambda: series(x**I, x))
    raises(PoleError, lambda: series(x**(I + 1), x))
    raises(PoleError, lambda: series(exp(I*log(x)), x))


def test_issue_14885():
    assert series(x**Rational(-3, 2)*exp(x), x, 0) == (x**Rational(-3, 2) + 1/sqrt(x) +
        sqrt(x)/2 + x**Rational(3, 2)/6 + x**Rational(5, 2)/24 + x**Rational(7, 2)/120 +
        x**Rational(9, 2)/720 + x**Rational(11, 2)/5040 + O(x**6))


def test_issue_15539():
    assert series(atan(x), x, -oo) == (-1/(5*x**5) + 1/(3*x**3) - 1/x - pi/2
        + O(x**(-6), (x, -oo)))
    assert series(atan(x), x, oo) == (-1/(5*x**5) + 1/(3*x**3) - 1/x + pi/2
        + O(x**(-6), (x, oo)))


def test_issue_7259():
    assert series(LambertW(x), x) == x - x**2 + 3*x**3/2 - 8*x**4/3 + 125*x**5/24 + O(x**6)
    assert series(LambertW(x**2), x, n=8) == x**2 - x**4 + 3*x**6/2 + O(x**8)
    assert series(LambertW(sin(x)), x, n=4) == x - x**2 + 4*x**3/3 + O(x**4)

def test_issue_11884():
    assert cos(x).series(x, 1, n=1) == cos(1) + O(x - 1, (x, 1))


def test_issue_18008():
    y = x*(1 + x*(1 - x))/((1 + x*(1 - x)) - (1 - x)*(1 - x))
    assert y.series(x, oo, n=4) == -9/(32*x**3) - 3/(16*x**2) - 1/(8*x) + S(1)/4 + x/2 + \
        O(x**(-4), (x, oo))


def test_issue_18842():
    f = log(x/(1 - x))
    assert f.series(x, 0.491, n=1).removeO().nsimplify() ==  \
        -S(180019443780011)/5000000000000000


def test_issue_19534():
    dt = symbols('dt', real=True)
    expr = 16*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0)/45 + \
            49*dt*(-0.049335189898860408029*dt*(2.0*dt + 1.0) + \
            0.29601113939316244817*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) - \
            0.12564355335492979587*dt*(0.074074074074074074074*dt*(2.0*dt + 1.0) + \
            0.2962962962962962963*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) + \
            0.96296296296296296296*dt + 1.0) + 0.051640768506639183825*dt + \
            dt*(1/2 - sqrt(21)/14) + 1.0)/180 + 49*dt*(-0.23637909581542530626*dt*(2.0*dt + 1.0) - \
            0.74817562366625959291*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) + \
            0.88085458023927036857*dt*(0.074074074074074074074*dt*(2.0*dt + 1.0) + \
            0.2962962962962962963*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) + \
            0.96296296296296296296*dt + 1.0) + \
            2.1165151389911680013*dt*(-0.049335189898860408029*dt*(2.0*dt + 1.0) + \
            0.29601113939316244817*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) - \
            0.12564355335492979587*dt*(0.074074074074074074074*dt*(2.0*dt + 1.0) + \
            0.2962962962962962963*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) + \
            0.96296296296296296296*dt + 1.0) + 0.22431393315265061193*dt + 1.0) - \
            1.1854881643947648988*dt + dt*(sqrt(21)/14 + 1/2) + 1.0)/180 + \
            dt*(0.66666666666666666667*dt*(2.0*dt + 1.0) + \
            6.0173399699313066769*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) - \
            4.1117044797036320069*dt*(0.074074074074074074074*dt*(2.0*dt + 1.0) + \
            0.2962962962962962963*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) + \
            0.96296296296296296296*dt + 1.0) - \
            7.0189140975801991157*dt*(-0.049335189898860408029*dt*(2.0*dt + 1.0) + \
            0.29601113939316244817*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) - \
            0.12564355335492979587*dt*(0.074074074074074074074*dt*(2.0*dt + 1.0) + \
            0.2962962962962962963*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) + \
            0.96296296296296296296*dt + 1.0) + 0.22431393315265061193*dt + 1.0) + \
            0.94010945196161777522*dt*(-0.23637909581542530626*dt*(2.0*dt + 1.0) - \
            0.74817562366625959291*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) + \
            0.88085458023927036857*dt*(0.074074074074074074074*dt*(2.0*dt + 1.0) + \
            0.2962962962962962963*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) + \
            0.96296296296296296296*dt + 1.0) + \
            2.1165151389911680013*dt*(-0.049335189898860408029*dt*(2.0*dt + 1.0) + \
            0.29601113939316244817*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) - \
            0.12564355335492979587*dt*(0.074074074074074074074*dt*(2.0*dt + 1.0) + \
            0.2962962962962962963*dt*(0.125*dt*(2.0*dt + 1.0) + 0.875*dt + 1.0) + \
            0.96296296296296296296*dt + 1.0) + 0.22431393315265061193*dt + 1.0) - \
            0.35816132904077632692*dt + 1.0) + 5.5065024887242400038*dt + 1.0)/20 + dt/20 + 1

    assert N(expr.series(dt, 0, 8), 20) == (
            - Float('0.00092592592592592596126289', precision=70) * dt**7
            + Float('0.0027777777777777783174695', precision=70) * dt**6
            + Float('0.016666666666666656027029', precision=70) * dt**5
            + Float('0.083333333333333300951828', precision=70) * dt**4
            + Float('0.33333333333333337034077', precision=70) * dt**3
            + Float('1.0', precision=70) * dt**2
            + Float('1.0', precision=70) * dt
            + Float('1.0', precision=70)
        )


def test_issue_11407():
    a, b, c, x = symbols('a b c x')
    assert series(sqrt(a + b + c*x), x, 0, 1) == sqrt(a + b) + O(x)
    assert series(sqrt(a + b + c + c*x), x, 0, 1) == sqrt(a + b + c) + O(x)


def test_issue_14037():
    assert (sin(x**50)/x**51).series(x, n=0) == 1/x + O(1, x)


def test_issue_20551():
    expr = (exp(x)/x).series(x, n=None)
    terms = [ next(expr) for i in range(3) ]
    assert terms == [1/x, 1, x/2]


def test_issue_20697():
    p_0, p_1, p_2, p_3, b_0, b_1, b_2 = symbols('p_0 p_1 p_2 p_3 b_0 b_1 b_2')
    Q = (p_0 + (p_1 + (p_2 + p_3/y)/y)/y)/(1 + ((p_3/(b_0*y) + (b_0*p_2\
        - b_1*p_3)/b_0**2)/y + (b_0**2*p_1 - b_0*b_1*p_2 - p_3*(b_0*b_2\
        - b_1**2))/b_0**3)/y)
    assert Q.series(y, n=3).ratsimp() == b_2*y**2 + b_1*y + b_0 + O(y**3)


def test_issue_21245():
    fi = (1 + sqrt(5))/2
    assert (1/(1 - x - x**2)).series(x, 1/fi, 1).factor() == \
        (-37*sqrt(5) - 83 + 13*sqrt(5)*x + 29*x + O((x - 2/(1 + sqrt(5)))**2, (x\
            , 2/(1 + sqrt(5)))))/((2*sqrt(5) + 5)**2*(x + sqrt(5)*x - 2))



def test_issue_21938():
    expr = sin(1/x + exp(-x)) - sin(1/x)
    assert expr.series(x, oo) == (1/(24*x**4) - 1/(2*x**2) + 1 + O(x**(-6), (x, oo)))*exp(-x)


def test_issue_23432():
    expr = 1/sqrt(1 - x**2)
    result = expr.series(x, 0.5)
    assert result.is_Add and len(result.args) == 7


def test_issue_23727():
    res = series(sqrt(1 - x**2), x, 0.1)
    assert res.is_Add == True


def test_issue_24266():
    #type1: exp(f(x))
    assert (exp(-I*pi*(2*x+1))).series(x, 0, 3) == -1 + 2*I*pi*x + 2*pi**2*x**2 + O(x**3)
    assert (exp(-I*pi*(2*x+1))*gamma(1+x)).series(x, 0, 3) == -1 + x*(EulerGamma + 2*I*pi) + \
        x**2*(-EulerGamma**2/2 + 23*pi**2/12 - 2*EulerGamma*I*pi) + O(x**3)

    #type2: c**f(x)
    assert ((2*I)**(-I*pi*(2*x+1))).series(x, 0, 2) == exp(pi**2/2 - I*pi*log(2)) + \
          x*(pi**2*exp(pi**2/2 - I*pi*log(2)) - 2*I*pi*exp(pi**2/2 - I*pi*log(2))*log(2)) + O(x**2)
    assert ((2)**(-I*pi*(2*x+1))).series(x, 0, 2) == exp(-I*pi*log(2)) - 2*I*pi*x*exp(-I*pi*log(2))*log(2) + O(x**2)

    #type3: f(y)**g(x)
    assert ((y)**(I*pi*(2*x+1))).series(x, 0, 2) == exp(I*pi*log(y)) + 2*I*pi*x*exp(I*pi*log(y))*log(y) + O(x**2)
    assert ((I*y)**(I*pi*(2*x+1))).series(x, 0, 2) == exp(I*pi*log(I*y)) + 2*I*pi*x*exp(I*pi*log(I*y))*log(I*y) + O(x**2)


def test_issue_26856():
    raises(ValueError, lambda: (2**x).series(x, oo, -1))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\indexes\categorical\test_indexing.py ===
import numpy as np
import pytest

from pandas.errors import InvalidIndexError

import pandas as pd
from pandas import (
    CategoricalIndex,
    Index,
    IntervalIndex,
    Timestamp,
)
import pandas._testing as tm


class TestTake:
    def test_take_fill_value(self):
        # GH 12631

        # numeric category
        idx = CategoricalIndex([1, 2, 3], name="xxx")
        result = idx.take(np.array([1, 0, -1]))
        expected = CategoricalIndex([2, 1, 3], name="xxx")
        tm.assert_index_equal(result, expected)
        tm.assert_categorical_equal(result.values, expected.values)

        # fill_value
        result = idx.take(np.array([1, 0, -1]), fill_value=True)
        expected = CategoricalIndex([2, 1, np.nan], categories=[1, 2, 3], name="xxx")
        tm.assert_index_equal(result, expected)
        tm.assert_categorical_equal(result.values, expected.values)

        # allow_fill=False
        result = idx.take(np.array([1, 0, -1]), allow_fill=False, fill_value=True)
        expected = CategoricalIndex([2, 1, 3], name="xxx")
        tm.assert_index_equal(result, expected)
        tm.assert_categorical_equal(result.values, expected.values)

        # object category
        idx = CategoricalIndex(
            list("CBA"), categories=list("ABC"), ordered=True, name="xxx"
        )
        result = idx.take(np.array([1, 0, -1]))
        expected = CategoricalIndex(
            list("BCA"), categories=list("ABC"), ordered=True, name="xxx"
        )
        tm.assert_index_equal(result, expected)
        tm.assert_categorical_equal(result.values, expected.values)

        # fill_value
        result = idx.take(np.array([1, 0, -1]), fill_value=True)
        expected = CategoricalIndex(
            ["B", "C", np.nan], categories=list("ABC"), ordered=True, name="xxx"
        )
        tm.assert_index_equal(result, expected)
        tm.assert_categorical_equal(result.values, expected.values)

        # allow_fill=False
        result = idx.take(np.array([1, 0, -1]), allow_fill=False, fill_value=True)
        expected = CategoricalIndex(
            list("BCA"), categories=list("ABC"), ordered=True, name="xxx"
        )
        tm.assert_index_equal(result, expected)
        tm.assert_categorical_equal(result.values, expected.values)

        msg = (
            "When allow_fill=True and fill_value is not None, "
            "all indices must be >= -1"
        )
        with pytest.raises(ValueError, match=msg):
            idx.take(np.array([1, 0, -2]), fill_value=True)
        with pytest.raises(ValueError, match=msg):
            idx.take(np.array([1, 0, -5]), fill_value=True)

        msg = "index -5 is out of bounds for (axis 0 with )?size 3"
        with pytest.raises(IndexError, match=msg):
            idx.take(np.array([1, -5]))

    def test_take_fill_value_datetime(self):
        # datetime category
        idx = pd.DatetimeIndex(["2011-01-01", "2011-02-01", "2011-03-01"], name="xxx")
        idx = CategoricalIndex(idx)
        result = idx.take(np.array([1, 0, -1]))
        expected = pd.DatetimeIndex(
            ["2011-02-01", "2011-01-01", "2011-03-01"], name="xxx"
        )
        expected = CategoricalIndex(expected)
        tm.assert_index_equal(result, expected)

        # fill_value
        result = idx.take(np.array([1, 0, -1]), fill_value=True)
        expected = pd.DatetimeIndex(["2011-02-01", "2011-01-01", "NaT"], name="xxx")
        exp_cats = pd.DatetimeIndex(["2011-01-01", "2011-02-01", "2011-03-01"])
        expected = CategoricalIndex(expected, categories=exp_cats)
        tm.assert_index_equal(result, expected)

        # allow_fill=False
        result = idx.take(np.array([1, 0, -1]), allow_fill=False, fill_value=True)
        expected = pd.DatetimeIndex(
            ["2011-02-01", "2011-01-01", "2011-03-01"], name="xxx"
        )
        expected = CategoricalIndex(expected)
        tm.assert_index_equal(result, expected)

        msg = (
            "When allow_fill=True and fill_value is not None, "
            "all indices must be >= -1"
        )
        with pytest.raises(ValueError, match=msg):
            idx.take(np.array([1, 0, -2]), fill_value=True)
        with pytest.raises(ValueError, match=msg):
            idx.take(np.array([1, 0, -5]), fill_value=True)

        msg = "index -5 is out of bounds for (axis 0 with )?size 3"
        with pytest.raises(IndexError, match=msg):
            idx.take(np.array([1, -5]))

    def test_take_invalid_kwargs(self):
        idx = CategoricalIndex([1, 2, 3], name="foo")
        indices = [1, 0, -1]

        msg = r"take\(\) got an unexpected keyword argument 'foo'"
        with pytest.raises(TypeError, match=msg):
            idx.take(indices, foo=2)

        msg = "the 'out' parameter is not supported"
        with pytest.raises(ValueError, match=msg):
            idx.take(indices, out=indices)

        msg = "the 'mode' parameter is not supported"
        with pytest.raises(ValueError, match=msg):
            idx.take(indices, mode="clip")


class TestGetLoc:
    def test_get_loc(self):
        # GH 12531
        cidx1 = CategoricalIndex(list("abcde"), categories=list("edabc"))
        idx1 = Index(list("abcde"))
        assert cidx1.get_loc("a") == idx1.get_loc("a")
        assert cidx1.get_loc("e") == idx1.get_loc("e")

        for i in [cidx1, idx1]:
            with pytest.raises(KeyError, match="'NOT-EXIST'"):
                i.get_loc("NOT-EXIST")

        # non-unique
        cidx2 = CategoricalIndex(list("aacded"), categories=list("edabc"))
        idx2 = Index(list("aacded"))

        # results in bool array
        res = cidx2.get_loc("d")
        tm.assert_numpy_array_equal(res, idx2.get_loc("d"))
        tm.assert_numpy_array_equal(
            res, np.array([False, False, False, True, False, True])
        )
        # unique element results in scalar
        res = cidx2.get_loc("e")
        assert res == idx2.get_loc("e")
        assert res == 4

        for i in [cidx2, idx2]:
            with pytest.raises(KeyError, match="'NOT-EXIST'"):
                i.get_loc("NOT-EXIST")

        # non-unique, sliceable
        cidx3 = CategoricalIndex(list("aabbb"), categories=list("abc"))
        idx3 = Index(list("aabbb"))

        # results in slice
        res = cidx3.get_loc("a")
        assert res == idx3.get_loc("a")
        assert res == slice(0, 2, None)

        res = cidx3.get_loc("b")
        assert res == idx3.get_loc("b")
        assert res == slice(2, 5, None)

        for i in [cidx3, idx3]:
            with pytest.raises(KeyError, match="'c'"):
                i.get_loc("c")

    def test_get_loc_unique(self):
        cidx = CategoricalIndex(list("abc"))
        result = cidx.get_loc("b")
        assert result == 1

    def test_get_loc_monotonic_nonunique(self):
        cidx = CategoricalIndex(list("abbc"))
        result = cidx.get_loc("b")
        expected = slice(1, 3, None)
        assert result == expected

    def test_get_loc_nonmonotonic_nonunique(self):
        cidx = CategoricalIndex(list("abcb"))
        result = cidx.get_loc("b")
        expected = np.array([False, True, False, True], dtype=bool)
        tm.assert_numpy_array_equal(result, expected)

    def test_get_loc_nan(self):
        # GH#41933
        ci = CategoricalIndex(["A", "B", np.nan])
        res = ci.get_loc(np.nan)

        assert res == 2


class TestGetIndexer:
    def test_get_indexer_base(self):
        # Determined by cat ordering.
        idx = CategoricalIndex(list("cab"), categories=list("cab"))
        expected = np.arange(len(idx), dtype=np.intp)

        actual = idx.get_indexer(idx)
        tm.assert_numpy_array_equal(expected, actual)

        with pytest.raises(ValueError, match="Invalid fill method"):
            idx.get_indexer(idx, method="invalid")

    def test_get_indexer_requires_unique(self):
        ci = CategoricalIndex(list("aabbca"), categories=list("cab"), ordered=False)
        oidx = Index(np.array(ci))

        msg = "Reindexing only valid with uniquely valued Index objects"

        for n in [1, 2, 5, len(ci)]:
            finder = oidx[np.random.default_rng(2).integers(0, len(ci), size=n)]

            with pytest.raises(InvalidIndexError, match=msg):
                ci.get_indexer(finder)

        # see gh-17323
        #
        # Even when indexer is equal to the
        # members in the index, we should
        # respect duplicates instead of taking
        # the fast-track path.
        for finder in [list("aabbca"), list("aababca")]:
            with pytest.raises(InvalidIndexError, match=msg):
                ci.get_indexer(finder)

    def test_get_indexer_non_unique(self):
        idx1 = CategoricalIndex(list("aabcde"), categories=list("edabc"))
        idx2 = CategoricalIndex(list("abf"))

        for indexer in [idx2, list("abf"), Index(list("abf"))]:
            msg = "Reindexing only valid with uniquely valued Index objects"
            with pytest.raises(InvalidIndexError, match=msg):
                idx1.get_indexer(indexer)

            r1, _ = idx1.get_indexer_non_unique(indexer)
            expected = np.array([0, 1, 2, -1], dtype=np.intp)
            tm.assert_almost_equal(r1, expected)

    def test_get_indexer_method(self):
        idx1 = CategoricalIndex(list("aabcde"), categories=list("edabc"))
        idx2 = CategoricalIndex(list("abf"))

        msg = "method pad not yet implemented for CategoricalIndex"
        with pytest.raises(NotImplementedError, match=msg):
            idx2.get_indexer(idx1, method="pad")
        msg = "method backfill not yet implemented for CategoricalIndex"
        with pytest.raises(NotImplementedError, match=msg):
            idx2.get_indexer(idx1, method="backfill")

        msg = "method nearest not yet implemented for CategoricalIndex"
        with pytest.raises(NotImplementedError, match=msg):
            idx2.get_indexer(idx1, method="nearest")

    def test_get_indexer_array(self):
        arr = np.array(
            [Timestamp("1999-12-31 00:00:00"), Timestamp("2000-12-31 00:00:00")],
            dtype=object,
        )
        cats = [Timestamp("1999-12-31 00:00:00"), Timestamp("2000-12-31 00:00:00")]
        ci = CategoricalIndex(cats, categories=cats, ordered=False, dtype="category")
        result = ci.get_indexer(arr)
        expected = np.array([0, 1], dtype="intp")
        tm.assert_numpy_array_equal(result, expected)

    def test_get_indexer_same_categories_same_order(self):
        ci = CategoricalIndex(["a", "b"], categories=["a", "b"])

        result = ci.get_indexer(CategoricalIndex(["b", "b"], categories=["a", "b"]))
        expected = np.array([1, 1], dtype="intp")
        tm.assert_numpy_array_equal(result, expected)

    def test_get_indexer_same_categories_different_order(self):
        # https://github.com/pandas-dev/pandas/issues/19551
        ci = CategoricalIndex(["a", "b"], categories=["a", "b"])

        result = ci.get_indexer(CategoricalIndex(["b", "b"], categories=["b", "a"]))
        expected = np.array([1, 1], dtype="intp")
        tm.assert_numpy_array_equal(result, expected)

    def test_get_indexer_nans_in_index_and_target(self):
        # GH 45361
        ci = CategoricalIndex([1, 2, np.nan, 3])
        other1 = [2, 3, 4, np.nan]
        res1 = ci.get_indexer(other1)
        expected1 = np.array([1, 3, -1, 2], dtype=np.intp)
        tm.assert_numpy_array_equal(res1, expected1)
        other2 = [1, 4, 2, 3]
        res2 = ci.get_indexer(other2)
        expected2 = np.array([0, -1, 1, 3], dtype=np.intp)
        tm.assert_numpy_array_equal(res2, expected2)


class TestWhere:
    def test_where(self, listlike_box):
        klass = listlike_box

        i = CategoricalIndex(list("aabbca"), categories=list("cab"), ordered=False)
        cond = [True] * len(i)
        expected = i
        result = i.where(klass(cond))
        tm.assert_index_equal(result, expected)

        cond = [False] + [True] * (len(i) - 1)
        expected = CategoricalIndex([np.nan] + i[1:].tolist(), categories=i.categories)
        result = i.where(klass(cond))
        tm.assert_index_equal(result, expected)

    def test_where_non_categories(self):
        ci = CategoricalIndex(["a", "b", "c", "d"])
        mask = np.array([True, False, True, False])

        result = ci.where(mask, 2)
        expected = Index(["a", 2, "c", 2], dtype=object)
        tm.assert_index_equal(result, expected)

        msg = "Cannot setitem on a Categorical with a new category"
        with pytest.raises(TypeError, match=msg):
            # Test the Categorical method directly
            ci._data._where(mask, 2)


class TestContains:
    def test_contains(self):
        ci = CategoricalIndex(list("aabbca"), categories=list("cabdef"), ordered=False)

        assert "a" in ci
        assert "z" not in ci
        assert "e" not in ci
        assert np.nan not in ci

        # assert codes NOT in index
        assert 0 not in ci
        assert 1 not in ci

    def test_contains_nan(self):
        ci = CategoricalIndex(list("aabbca") + [np.nan], categories=list("cabdef"))
        assert np.nan in ci

    @pytest.mark.parametrize("unwrap", [True, False])
    def test_contains_na_dtype(self, unwrap):
        dti = pd.date_range("2016-01-01", periods=100).insert(0, pd.NaT)
        pi = dti.to_period("D")
        tdi = dti - dti[-1]
        ci = CategoricalIndex(dti)

        obj = ci
        if unwrap:
            obj = ci._data

        assert np.nan in obj
        assert None in obj
        assert pd.NaT in obj
        assert np.datetime64("NaT") in obj
        assert np.timedelta64("NaT") not in obj

        obj2 = CategoricalIndex(tdi)
        if unwrap:
            obj2 = obj2._data

        assert np.nan in obj2
        assert None in obj2
        assert pd.NaT in obj2
        assert np.datetime64("NaT") not in obj2
        assert np.timedelta64("NaT") in obj2

        obj3 = CategoricalIndex(pi)
        if unwrap:
            obj3 = obj3._data

        assert np.nan in obj3
        assert None in obj3
        assert pd.NaT in obj3
        assert np.datetime64("NaT") not in obj3
        assert np.timedelta64("NaT") not in obj3

    @pytest.mark.parametrize(
        "item, expected",
        [
            (pd.Interval(0, 1), True),
            (1.5, True),
            (pd.Interval(0.5, 1.5), False),
            ("a", False),
            (Timestamp(1), False),
            (pd.Timedelta(1), False),
        ],
        ids=str,
    )
    def test_contains_interval(self, item, expected):
        # GH 23705
        ci = CategoricalIndex(IntervalIndex.from_breaks(range(3)))
        result = item in ci
        assert result is expected

    def test_contains_list(self):
        # GH#21729
        idx = CategoricalIndex([1, 2, 3])

        assert "a" not in idx

        with pytest.raises(TypeError, match="unhashable type"):
            ["a"] in idx

        with pytest.raises(TypeError, match="unhashable type"):
            ["a", "b"] in idx