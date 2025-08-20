
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\exceptions.py ===
"""
Exceptions and Warnings
=======================

General exceptions used by NumPy.  Note that some exceptions may be module
specific, such as linear algebra errors.

.. versionadded:: NumPy 1.25

    The exceptions module is new in NumPy 1.25.  Older exceptions remain
    available through the main NumPy namespace for compatibility.

.. currentmodule:: numpy.exceptions

Warnings
--------
.. autosummary::
   :toctree: generated/

   ComplexWarning             Given when converting complex to real.
   VisibleDeprecationWarning  Same as a DeprecationWarning, but more visible.
   RankWarning                Issued when the design matrix is rank deficient.

Exceptions
----------
.. autosummary::
   :toctree: generated/

    AxisError          Given when an axis was invalid.
    DTypePromotionError   Given when no common dtype could be found.
    TooHardError       Error specific to `numpy.shares_memory`.

"""


__all__ = [
    "ComplexWarning", "VisibleDeprecationWarning", "ModuleDeprecationWarning",
    "TooHardError", "AxisError", "DTypePromotionError"]


# Disallow reloading this module so as to preserve the identities of the
# classes defined here.
if '_is_loaded' in globals():
    raise RuntimeError('Reloading numpy._globals is not allowed')
_is_loaded = True


class ComplexWarning(RuntimeWarning):
    """
    The warning raised when casting a complex dtype to a real dtype.

    As implemented, casting a complex number to a real discards its imaginary
    part, but this behavior may not be what the user actually wants.

    """
    pass


class ModuleDeprecationWarning(DeprecationWarning):
    """Module deprecation warning.

    .. warning::

        This warning should not be used, since nose testing is not relevant
        anymore.

    The nose tester turns ordinary Deprecation warnings into test failures.
    That makes it hard to deprecate whole modules, because they get
    imported by default. So this is a special Deprecation warning that the
    nose tester will let pass without making tests fail.

    """
    pass


class VisibleDeprecationWarning(UserWarning):
    """Visible deprecation warning.

    By default, python will not show deprecation warnings, so this class
    can be used when a very visible warning is helpful, for example because
    the usage is most likely a user bug.

    """
    pass


class RankWarning(RuntimeWarning):
    """Matrix rank warning.

    Issued by polynomial functions when the design matrix is rank deficient.

    """
    pass


# Exception used in shares_memory()
class TooHardError(RuntimeError):
    """``max_work`` was exceeded.

    This is raised whenever the maximum number of candidate solutions
    to consider specified by the ``max_work`` parameter is exceeded.
    Assigning a finite number to ``max_work`` may have caused the operation
    to fail.

    """
    pass


class AxisError(ValueError, IndexError):
    """Axis supplied was invalid.

    This is raised whenever an ``axis`` parameter is specified that is larger
    than the number of array dimensions.
    For compatibility with code written against older numpy versions, which
    raised a mixture of :exc:`ValueError` and :exc:`IndexError` for this
    situation, this exception subclasses both to ensure that
    ``except ValueError`` and ``except IndexError`` statements continue
    to catch ``AxisError``.

    Parameters
    ----------
    axis : int or str
        The out of bounds axis or a custom exception message.
        If an axis is provided, then `ndim` should be specified as well.
    ndim : int, optional
        The number of array dimensions.
    msg_prefix : str, optional
        A prefix for the exception message.

    Attributes
    ----------
    axis : int, optional
        The out of bounds axis or ``None`` if a custom exception
        message was provided. This should be the axis as passed by
        the user, before any normalization to resolve negative indices.

        .. versionadded:: 1.22
    ndim : int, optional
        The number of array dimensions or ``None`` if a custom exception
        message was provided.

        .. versionadded:: 1.22


    Examples
    --------
    >>> import numpy as np
    >>> array_1d = np.arange(10)
    >>> np.cumsum(array_1d, axis=1)
    Traceback (most recent call last):
      ...
    numpy.exceptions.AxisError: axis 1 is out of bounds for array of dimension 1

    Negative axes are preserved:

    >>> np.cumsum(array_1d, axis=-2)
    Traceback (most recent call last):
      ...
    numpy.exceptions.AxisError: axis -2 is out of bounds for array of dimension 1

    The class constructor generally takes the axis and arrays'
    dimensionality as arguments:

    >>> print(np.exceptions.AxisError(2, 1, msg_prefix='error'))
    error: axis 2 is out of bounds for array of dimension 1

    Alternatively, a custom exception message can be passed:

    >>> print(np.exceptions.AxisError('Custom error message'))
    Custom error message

    """

    __slots__ = ("_msg", "axis", "ndim")

    def __init__(self, axis, ndim=None, msg_prefix=None):
        if ndim is msg_prefix is None:
            # single-argument form: directly set the error message
            self._msg = axis
            self.axis = None
            self.ndim = None
        else:
            self._msg = msg_prefix
            self.axis = axis
            self.ndim = ndim

    def __str__(self):
        axis = self.axis
        ndim = self.ndim

        if axis is ndim is None:
            return self._msg
        else:
            msg = f"axis {axis} is out of bounds for array of dimension {ndim}"
            if self._msg is not None:
                msg = f"{self._msg}: {msg}"
            return msg


class DTypePromotionError(TypeError):
    """Multiple DTypes could not be converted to a common one.

    This exception derives from ``TypeError`` and is raised whenever dtypes
    cannot be converted to a single common one.  This can be because they
    are of a different category/class or incompatible instances of the same
    one (see Examples).

    Notes
    -----
    Many functions will use promotion to find the correct result and
    implementation.  For these functions the error will typically be chained
    with a more specific error indicating that no implementation was found
    for the input dtypes.

    Typically promotion should be considered "invalid" between the dtypes of
    two arrays when `arr1 == arr2` can safely return all ``False`` because the
    dtypes are fundamentally different.

    Examples
    --------
    Datetimes and complex numbers are incompatible classes and cannot be
    promoted:

    >>> import numpy as np
    >>> np.result_type(np.dtype("M8[s]"), np.complex128)  # doctest: +IGNORE_EXCEPTION_DETAIL
    Traceback (most recent call last):
     ...
    DTypePromotionError: The DType <class 'numpy.dtype[datetime64]'> could not
    be promoted by <class 'numpy.dtype[complex128]'>. This means that no common
    DType exists for the given inputs. For example they cannot be stored in a
    single array unless the dtype is `object`. The full list of DTypes is:
    (<class 'numpy.dtype[datetime64]'>, <class 'numpy.dtype[complex128]'>)

    For example for structured dtypes, the structure can mismatch and the
    same ``DTypePromotionError`` is given when two structured dtypes with
    a mismatch in their number of fields is given:

    >>> dtype1 = np.dtype([("field1", np.float64), ("field2", np.int64)])
    >>> dtype2 = np.dtype([("field1", np.float64)])
    >>> np.promote_types(dtype1, dtype2)  # doctest: +IGNORE_EXCEPTION_DETAIL
    Traceback (most recent call last):
     ...
    DTypePromotionError: field names `('field1', 'field2')` and `('field1',)`
    mismatch.

    """  # noqa: E501
    pass

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\f2py\_src_pyf.py ===
import os
import re

# START OF CODE VENDORED FROM `numpy.distutils.from_template`
#############################################################
"""
process_file(filename)

  takes templated file .xxx.src and produces .xxx file where .xxx
  is .pyf .f90 or .f using the following template rules:

  '<..>' denotes a template.

  All function and subroutine blocks in a source file with names that
  contain '<..>' will be replicated according to the rules in '<..>'.

  The number of comma-separated words in '<..>' will determine the number of
  replicates.

  '<..>' may have two different forms, named and short. For example,

  named:
   <p=d,s,z,c> where anywhere inside a block '<p>' will be replaced with
   'd', 's', 'z', and 'c' for each replicate of the block.

   <_c>  is already defined: <_c=s,d,c,z>
   <_t>  is already defined: <_t=real,double precision,complex,double complex>

  short:
   <s,d,c,z>, a short form of the named, useful when no <p> appears inside
   a block.

  In general, '<..>' contains a comma separated list of arbitrary
  expressions. If these expression must contain a comma|leftarrow|rightarrow,
  then prepend the comma|leftarrow|rightarrow with a backslash.

  If an expression matches '\\<index>' then it will be replaced
  by <index>-th expression.

  Note that all '<..>' forms in a block must have the same number of
  comma-separated entries.

 Predefined named template rules:
  <prefix=s,d,c,z>
  <ftype=real,double precision,complex,double complex>
  <ftypereal=real,double precision,\\0,\\1>
  <ctype=float,double,complex_float,complex_double>
  <ctypereal=float,double,\\0,\\1>
"""

routine_start_re = re.compile(r'(\n|\A)((     (\$|\*))|)\s*(subroutine|function)\b', re.I)
routine_end_re = re.compile(r'\n\s*end\s*(subroutine|function)\b.*(\n|\Z)', re.I)
function_start_re = re.compile(r'\n     (\$|\*)\s*function\b', re.I)

def parse_structure(astr):
    """ Return a list of tuples for each function or subroutine each
    tuple is the start and end of a subroutine or function to be
    expanded.
    """

    spanlist = []
    ind = 0
    while True:
        m = routine_start_re.search(astr, ind)
        if m is None:
            break
        start = m.start()
        if function_start_re.match(astr, start, m.end()):
            while True:
                i = astr.rfind('\n', ind, start)
                if i == -1:
                    break
                start = i
                if astr[i:i + 7] != '\n     $':
                    break
        start += 1
        m = routine_end_re.search(astr, m.end())
        ind = end = (m and m.end() - 1) or len(astr)
        spanlist.append((start, end))
    return spanlist


template_re = re.compile(r"<\s*(\w[\w\d]*)\s*>")
named_re = re.compile(r"<\s*(\w[\w\d]*)\s*=\s*(.*?)\s*>")
list_re = re.compile(r"<\s*((.*?))\s*>")

def find_repl_patterns(astr):
    reps = named_re.findall(astr)
    names = {}
    for rep in reps:
        name = rep[0].strip() or unique_key(names)
        repl = rep[1].replace(r'\,', '@comma@')
        thelist = conv(repl)
        names[name] = thelist
    return names

def find_and_remove_repl_patterns(astr):
    names = find_repl_patterns(astr)
    astr = re.subn(named_re, '', astr)[0]
    return astr, names


item_re = re.compile(r"\A\\(?P<index>\d+)\Z")
def conv(astr):
    b = astr.split(',')
    l = [x.strip() for x in b]
    for i in range(len(l)):
        m = item_re.match(l[i])
        if m:
            j = int(m.group('index'))
            l[i] = l[j]
    return ','.join(l)

def unique_key(adict):
    """ Obtain a unique key given a dictionary."""
    allkeys = list(adict.keys())
    done = False
    n = 1
    while not done:
        newkey = f'__l{n}'
        if newkey in allkeys:
            n += 1
        else:
            done = True
    return newkey


template_name_re = re.compile(r'\A\s*(\w[\w\d]*)\s*\Z')
def expand_sub(substr, names):
    substr = substr.replace(r'\>', '@rightarrow@')
    substr = substr.replace(r'\<', '@leftarrow@')
    lnames = find_repl_patterns(substr)
    substr = named_re.sub(r"<\1>", substr)  # get rid of definition templates

    def listrepl(mobj):
        thelist = conv(mobj.group(1).replace(r'\,', '@comma@'))
        if template_name_re.match(thelist):
            return f"<{thelist}>"
        name = None
        for key in lnames.keys():    # see if list is already in dictionary
            if lnames[key] == thelist:
                name = key
        if name is None:      # this list is not in the dictionary yet
            name = unique_key(lnames)
            lnames[name] = thelist
        return f"<{name}>"

    # convert all lists to named templates
    # new names are constructed as needed
    substr = list_re.sub(listrepl, substr)

    numsubs = None
    base_rule = None
    rules = {}
    for r in template_re.findall(substr):
        if r not in rules:
            thelist = lnames.get(r, names.get(r, None))
            if thelist is None:
                raise ValueError(f'No replicates found for <{r}>')
            if r not in names and not thelist.startswith('_'):
                names[r] = thelist
            rule = [i.replace('@comma@', ',') for i in thelist.split(',')]
            num = len(rule)

            if numsubs is None:
                numsubs = num
                rules[r] = rule
                base_rule = r
            elif num == numsubs:
                rules[r] = rule
            else:
                rules_base_rule = ','.join(rules[base_rule])
                print("Mismatch in number of replacements "
                      f"(base <{base_rule}={rules_base_rule}>) "
                      f"for <{r}={thelist}>. Ignoring.")
    if not rules:
        return substr

    def namerepl(mobj):
        name = mobj.group(1)
        return rules.get(name, (k + 1) * [name])[k]

    newstr = ''
    for k in range(numsubs):
        newstr += template_re.sub(namerepl, substr) + '\n\n'

    newstr = newstr.replace('@rightarrow@', '>')
    newstr = newstr.replace('@leftarrow@', '<')
    return newstr

def process_str(allstr):
    newstr = allstr
    writestr = ''

    struct = parse_structure(newstr)

    oldend = 0
    names = {}
    names.update(_special_names)
    for sub in struct:
        cleanedstr, defs = find_and_remove_repl_patterns(newstr[oldend:sub[0]])
        writestr += cleanedstr
        names.update(defs)
        writestr += expand_sub(newstr[sub[0]:sub[1]], names)
        oldend = sub[1]
    writestr += newstr[oldend:]

    return writestr


include_src_re = re.compile(r"(\n|\A)\s*include\s*['\"](?P<name>[\w\d./\\]+\.src)['\"]", re.I)

def resolve_includes(source):
    d = os.path.dirname(source)
    with open(source) as fid:
        lines = []
        for line in fid:
            m = include_src_re.match(line)
            if m:
                fn = m.group('name')
                if not os.path.isabs(fn):
                    fn = os.path.join(d, fn)
                if os.path.isfile(fn):
                    lines.extend(resolve_includes(fn))
                else:
                    lines.append(line)
            else:
                lines.append(line)
    return lines

def process_file(source):
    lines = resolve_includes(source)
    return process_str(''.join(lines))


_special_names = find_repl_patterns('''
<_c=s,d,c,z>
<_t=real,double precision,complex,double complex>
<prefix=s,d,c,z>
<ftype=real,double precision,complex,double complex>
<ctype=float,double,complex_float,complex_double>
<ftypereal=real,double precision,\\0,\\1>
<ctypereal=float,double,\\0,\\1>
''')

# END OF CODE VENDORED FROM `numpy.distutils.from_template`
###########################################################

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\pygments\sphinxext.py ===
"""
    pygments.sphinxext
    ~~~~~~~~~~~~~~~~~~

    Sphinx extension to generate automatic documentation of lexers,
    formatters and filters.

    :copyright: Copyright 2006-2025 by the Pygments team, see AUTHORS.
    :license: BSD, see LICENSE for details.
"""

import sys

from docutils import nodes
from docutils.statemachine import ViewList
from docutils.parsers.rst import Directive
from sphinx.util.nodes import nested_parse_with_titles


MODULEDOC = '''
.. module:: %s

%s
%s
'''

LEXERDOC = '''
.. class:: %s

    :Short names: %s
    :Filenames:   %s
    :MIME types:  %s

    %s

    %s

'''

FMTERDOC = '''
.. class:: %s

    :Short names: %s
    :Filenames: %s

    %s

'''

FILTERDOC = '''
.. class:: %s

    :Name: %s

    %s

'''


class PygmentsDoc(Directive):
    """
    A directive to collect all lexers/formatters/filters and generate
    autoclass directives for them.
    """
    has_content = False
    required_arguments = 1
    optional_arguments = 0
    final_argument_whitespace = False
    option_spec = {}

    def run(self):
        self.filenames = set()
        if self.arguments[0] == 'lexers':
            out = self.document_lexers()
        elif self.arguments[0] == 'formatters':
            out = self.document_formatters()
        elif self.arguments[0] == 'filters':
            out = self.document_filters()
        elif self.arguments[0] == 'lexers_overview':
            out = self.document_lexers_overview()
        else:
            raise Exception('invalid argument for "pygmentsdoc" directive')
        node = nodes.compound()
        vl = ViewList(out.split('\n'), source='')
        nested_parse_with_titles(self.state, vl, node)
        for fn in self.filenames:
            self.state.document.settings.record_dependencies.add(fn)
        return node.children

    def document_lexers_overview(self):
        """Generate a tabular overview of all lexers.

        The columns are the lexer name, the extensions handled by this lexer
        (or "None"), the aliases and a link to the lexer class."""
        from pip._vendor.pygments.lexers._mapping import LEXERS
        from pip._vendor.pygments.lexers import find_lexer_class
        out = []

        table = []

        def format_link(name, url):
            if url:
                return f'`{name} <{url}>`_'
            return name

        for classname, data in sorted(LEXERS.items(), key=lambda x: x[1][1].lower()):
            lexer_cls = find_lexer_class(data[1])
            extensions = lexer_cls.filenames + lexer_cls.alias_filenames

            table.append({
                'name': format_link(data[1], lexer_cls.url),
                'extensions': ', '.join(extensions).replace('*', '\\*').replace('_', '\\') or 'None',
                'aliases': ', '.join(data[2]),
                'class': f'{data[0]}.{classname}'
            })

        column_names = ['name', 'extensions', 'aliases', 'class']
        column_lengths = [max([len(row[column]) for row in table if row[column]])
                          for column in column_names]

        def write_row(*columns):
            """Format a table row"""
            out = []
            for length, col in zip(column_lengths, columns):
                if col:
                    out.append(col.ljust(length))
                else:
                    out.append(' '*length)

            return ' '.join(out)

        def write_seperator():
            """Write a table separator row"""
            sep = ['='*c for c in column_lengths]
            return write_row(*sep)

        out.append(write_seperator())
        out.append(write_row('Name', 'Extension(s)', 'Short name(s)', 'Lexer class'))
        out.append(write_seperator())
        for row in table:
            out.append(write_row(
                row['name'],
                row['extensions'],
                row['aliases'],
                f':class:`~{row["class"]}`'))
        out.append(write_seperator())

        return '\n'.join(out)

    def document_lexers(self):
        from pip._vendor.pygments.lexers._mapping import LEXERS
        from pip._vendor import pygments
        import inspect
        import pathlib

        out = []
        modules = {}
        moduledocstrings = {}
        for classname, data in sorted(LEXERS.items(), key=lambda x: x[0]):
            module = data[0]
            mod = __import__(module, None, None, [classname])
            self.filenames.add(mod.__file__)
            cls = getattr(mod, classname)
            if not cls.__doc__:
                print(f"Warning: {classname} does not have a docstring.")
            docstring = cls.__doc__
            if isinstance(docstring, bytes):
                docstring = docstring.decode('utf8')

            example_file = getattr(cls, '_example', None)
            if example_file:
                p = pathlib.Path(inspect.getabsfile(pygments)).parent.parent /\
                    'tests' / 'examplefiles' / example_file
                content = p.read_text(encoding='utf-8')
                if not content:
                    raise Exception(
                        f"Empty example file '{example_file}' for lexer "
                        f"{classname}")

                if data[2]:
                    lexer_name = data[2][0]
                    docstring += '\n\n    .. admonition:: Example\n'
                    docstring += f'\n      .. code-block:: {lexer_name}\n\n'
                    for line in content.splitlines():
                        docstring += f'          {line}\n'

            if cls.version_added:
                version_line = f'.. versionadded:: {cls.version_added}'
            else:
                version_line = ''

            modules.setdefault(module, []).append((
                classname,
                ', '.join(data[2]) or 'None',
                ', '.join(data[3]).replace('*', '\\*').replace('_', '\\') or 'None',
                ', '.join(data[4]) or 'None',
                docstring,
                version_line))
            if module not in moduledocstrings:
                moddoc = mod.__doc__
                if isinstance(moddoc, bytes):
                    moddoc = moddoc.decode('utf8')
                moduledocstrings[module] = moddoc

        for module, lexers in sorted(modules.items(), key=lambda x: x[0]):
            if moduledocstrings[module] is None:
                raise Exception(f"Missing docstring for {module}")
            heading = moduledocstrings[module].splitlines()[4].strip().rstrip('.')
            out.append(MODULEDOC % (module, heading, '-'*len(heading)))
            for data in lexers:
                out.append(LEXERDOC % data)

        return ''.join(out)

    def document_formatters(self):
        from pip._vendor.pygments.formatters import FORMATTERS

        out = []
        for classname, data in sorted(FORMATTERS.items(), key=lambda x: x[0]):
            module = data[0]
            mod = __import__(module, None, None, [classname])
            self.filenames.add(mod.__file__)
            cls = getattr(mod, classname)
            docstring = cls.__doc__
            if isinstance(docstring, bytes):
                docstring = docstring.decode('utf8')
            heading = cls.__name__
            out.append(FMTERDOC % (heading, ', '.join(data[2]) or 'None',
                                   ', '.join(data[3]).replace('*', '\\*') or 'None',
                                   docstring))
        return ''.join(out)

    def document_filters(self):
        from pip._vendor.pygments.filters import FILTERS

        out = []
        for name, cls in FILTERS.items():
            self.filenames.add(sys.modules[cls.__module__].__file__)
            docstring = cls.__doc__
            if isinstance(docstring, bytes):
                docstring = docstring.decode('utf8')
            out.append(FILTERDOC % (cls.__name__, name, docstring))
        return ''.join(out)


def setup(app):
    app.add_directive('pygmentsdoc', PygmentsDoc)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\connectors\pyodbc.py ===
# connectors/pyodbc.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

from __future__ import annotations

import re
from types import ModuleType
import typing
from typing import Any
from typing import Dict
from typing import List
from typing import Optional
from typing import Tuple
from typing import Union
from urllib.parse import unquote_plus

from . import Connector
from .. import ExecutionContext
from .. import pool
from .. import util
from ..engine import ConnectArgsType
from ..engine import Connection
from ..engine import interfaces
from ..engine import URL
from ..sql.type_api import TypeEngine

if typing.TYPE_CHECKING:
    from ..engine.interfaces import IsolationLevel


class PyODBCConnector(Connector):
    driver = "pyodbc"

    # this is no longer False for pyodbc in general
    supports_sane_rowcount_returning = True
    supports_sane_multi_rowcount = False

    supports_native_decimal = True
    default_paramstyle = "named"

    fast_executemany = False

    # for non-DSN connections, this *may* be used to
    # hold the desired driver name
    pyodbc_driver_name: Optional[str] = None

    dbapi: ModuleType

    def __init__(self, use_setinputsizes: bool = False, **kw: Any):
        super().__init__(**kw)
        if use_setinputsizes:
            self.bind_typing = interfaces.BindTyping.SETINPUTSIZES

    @classmethod
    def import_dbapi(cls) -> ModuleType:
        return __import__("pyodbc")

    def create_connect_args(self, url: URL) -> ConnectArgsType:
        opts = url.translate_connect_args(username="user")
        opts.update(url.query)

        keys = opts

        query = url.query

        connect_args: Dict[str, Any] = {}
        connectors: List[str]

        for param in ("ansi", "unicode_results", "autocommit"):
            if param in keys:
                connect_args[param] = util.asbool(keys.pop(param))

        if "odbc_connect" in keys:
            connectors = [unquote_plus(keys.pop("odbc_connect"))]
        else:

            def check_quote(token: str) -> str:
                if ";" in str(token) or str(token).startswith("{"):
                    token = "{%s}" % token.replace("}", "}}")
                return token

            keys = {k: check_quote(v) for k, v in keys.items()}

            dsn_connection = "dsn" in keys or (
                "host" in keys and "database" not in keys
            )
            if dsn_connection:
                connectors = [
                    "dsn=%s" % (keys.pop("host", "") or keys.pop("dsn", ""))
                ]
            else:
                port = ""
                if "port" in keys and "port" not in query:
                    port = ",%d" % int(keys.pop("port"))

                connectors = []
                driver = keys.pop("driver", self.pyodbc_driver_name)
                if driver is None and keys:
                    # note if keys is empty, this is a totally blank URL
                    util.warn(
                        "No driver name specified; "
                        "this is expected by PyODBC when using "
                        "DSN-less connections"
                    )
                else:
                    connectors.append("DRIVER={%s}" % driver)

                connectors.extend(
                    [
                        "Server=%s%s" % (keys.pop("host", ""), port),
                        "Database=%s" % keys.pop("database", ""),
                    ]
                )

            user = keys.pop("user", None)
            if user:
                connectors.append("UID=%s" % user)
                pwd = keys.pop("password", "")
                if pwd:
                    connectors.append("PWD=%s" % pwd)
            else:
                authentication = keys.pop("authentication", None)
                if authentication:
                    connectors.append("Authentication=%s" % authentication)
                else:
                    connectors.append("Trusted_Connection=Yes")

            # if set to 'Yes', the ODBC layer will try to automagically
            # convert textual data from your database encoding to your
            # client encoding.  This should obviously be set to 'No' if
            # you query a cp1253 encoded database from a latin1 client...
            if "odbc_autotranslate" in keys:
                connectors.append(
                    "AutoTranslate=%s" % keys.pop("odbc_autotranslate")
                )

            connectors.extend(["%s=%s" % (k, v) for k, v in keys.items()])

        return ((";".join(connectors),), connect_args)

    def is_disconnect(
        self,
        e: Exception,
        connection: Optional[
            Union[pool.PoolProxiedConnection, interfaces.DBAPIConnection]
        ],
        cursor: Optional[interfaces.DBAPICursor],
    ) -> bool:
        if isinstance(e, self.dbapi.ProgrammingError):
            return "The cursor's connection has been closed." in str(
                e
            ) or "Attempt to use a closed connection." in str(e)
        else:
            return False

    def _dbapi_version(self) -> interfaces.VersionInfoType:
        if not self.dbapi:
            return ()
        return self._parse_dbapi_version(self.dbapi.version)

    def _parse_dbapi_version(self, vers: str) -> interfaces.VersionInfoType:
        m = re.match(r"(?:py.*-)?([\d\.]+)(?:-(\w+))?", vers)
        if not m:
            return ()
        vers_tuple: interfaces.VersionInfoType = tuple(
            [int(x) for x in m.group(1).split(".")]
        )
        if m.group(2):
            vers_tuple += (m.group(2),)
        return vers_tuple

    def _get_server_version_info(
        self, connection: Connection
    ) -> interfaces.VersionInfoType:
        # NOTE: this function is not reliable, particularly when
        # freetds is in use.   Implement database-specific server version
        # queries.
        dbapi_con = connection.connection.dbapi_connection
        version: Tuple[Union[int, str], ...] = ()
        r = re.compile(r"[.\-]")
        for n in r.split(dbapi_con.getinfo(self.dbapi.SQL_DBMS_VER)):  # type: ignore[union-attr]  # noqa: E501
            try:
                version += (int(n),)
            except ValueError:
                pass
        return tuple(version)

    def do_set_input_sizes(
        self,
        cursor: interfaces.DBAPICursor,
        list_of_tuples: List[Tuple[str, Any, TypeEngine[Any]]],
        context: ExecutionContext,
    ) -> None:
        # the rules for these types seems a little strange, as you can pass
        # non-tuples as well as tuples, however it seems to assume "0"
        # for the subsequent values if you don't pass a tuple which fails
        # for types such as pyodbc.SQL_WLONGVARCHAR, which is the datatype
        # that ticket #5649 is targeting.

        # NOTE: as of #6058, this won't be called if the use_setinputsizes
        # parameter were not passed to the dialect, or if no types were
        # specified in list_of_tuples

        # as of #8177 for 2.0 we assume use_setinputsizes=True and only
        # omit the setinputsizes calls for .executemany() with
        # fast_executemany=True

        if (
            context.execute_style is interfaces.ExecuteStyle.EXECUTEMANY
            and self.fast_executemany
        ):
            return

        cursor.setinputsizes(
            [
                (
                    (dbtype, None, None)
                    if not isinstance(dbtype, tuple)
                    else dbtype
                )
                for key, dbtype, sqltype in list_of_tuples
            ]
        )

    def get_isolation_level_values(
        self, dbapi_conn: interfaces.DBAPIConnection
    ) -> List[IsolationLevel]:
        return [*super().get_isolation_level_values(dbapi_conn), "AUTOCOMMIT"]

    def set_isolation_level(
        self,
        dbapi_connection: interfaces.DBAPIConnection,
        level: IsolationLevel,
    ) -> None:
        # adjust for ConnectionFairy being present
        # allows attribute set e.g. "connection.autocommit = True"
        # to work properly

        if level == "AUTOCOMMIT":
            dbapi_connection.autocommit = True
        else:
            dbapi_connection.autocommit = False
            super().set_isolation_level(dbapi_connection, level)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\mpmath\calculus\approximation.py ===
from ..libmp.backend import xrange
from .calculus import defun

#----------------------------------------------------------------------------#
#                              Approximation methods                         #
#----------------------------------------------------------------------------#

# The Chebyshev approximation formula is given at:
# http://mathworld.wolfram.com/ChebyshevApproximationFormula.html

# The only major changes in the following code is that we return the
# expanded polynomial coefficients instead of Chebyshev coefficients,
# and that we automatically transform [a,b] -> [-1,1] and back
# for convenience.

# Coefficient in Chebyshev approximation
def chebcoeff(ctx,f,a,b,j,N):
    s = ctx.mpf(0)
    h = ctx.mpf(0.5)
    for k in range(1, N+1):
        t = ctx.cospi((k-h)/N)
        s += f(t*(b-a)*h + (b+a)*h) * ctx.cospi(j*(k-h)/N)
    return 2*s/N

# Generate Chebyshev polynomials T_n(ax+b) in expanded form
def chebT(ctx, a=1, b=0):
    Tb = [1]
    yield Tb
    Ta = [b, a]
    while 1:
        yield Ta
        # Recurrence: T[n+1](ax+b) = 2*(ax+b)*T[n](ax+b) - T[n-1](ax+b)
        Tmp = [0] + [2*a*t for t in Ta]
        for i, c in enumerate(Ta): Tmp[i] += 2*b*c
        for i, c in enumerate(Tb): Tmp[i] -= c
        Ta, Tb = Tmp, Ta

@defun
def chebyfit(ctx, f, interval, N, error=False):
    r"""
    Computes a polynomial of degree `N-1` that approximates the
    given function `f` on the interval `[a, b]`. With ``error=True``,
    :func:`~mpmath.chebyfit` also returns an accurate estimate of the
    maximum absolute error; that is, the maximum value of
    `|f(x) - P(x)|` for `x \in [a, b]`.

    :func:`~mpmath.chebyfit` uses the Chebyshev approximation formula,
    which gives a nearly optimal solution: that is, the maximum
    error of the approximating polynomial is very close to
    the smallest possible for any polynomial of the same degree.

    Chebyshev approximation is very useful if one needs repeated
    evaluation of an expensive function, such as function defined
    implicitly by an integral or a differential equation. (For
    example, it could be used to turn a slow mpmath function
    into a fast machine-precision version of the same.)

    **Examples**

    Here we use :func:`~mpmath.chebyfit` to generate a low-degree approximation
    of `f(x) = \cos(x)`, valid on the interval `[1, 2]`::

        >>> from mpmath import *
        >>> mp.dps = 15; mp.pretty = True
        >>> poly, err = chebyfit(cos, [1, 2], 5, error=True)
        >>> nprint(poly)
        [0.00291682, 0.146166, -0.732491, 0.174141, 0.949553]
        >>> nprint(err, 12)
        1.61351758081e-5

    The polynomial can be evaluated using ``polyval``::

        >>> nprint(polyval(poly, 1.6), 12)
        -0.0291858904138
        >>> nprint(cos(1.6), 12)
        -0.0291995223013

    Sampling the true error at 1000 points shows that the error
    estimate generated by ``chebyfit`` is remarkably good::

        >>> error = lambda x: abs(cos(x) - polyval(poly, x))
        >>> nprint(max([error(1+n/1000.) for n in range(1000)]), 12)
        1.61349954245e-5

    **Choice of degree**

    The degree `N` can be set arbitrarily high, to obtain an
    arbitrarily good approximation. As a rule of thumb, an
    `N`-term Chebyshev approximation is good to `N/(b-a)` decimal
    places on a unit interval (although this depends on how
    well-behaved `f` is). The cost grows accordingly: ``chebyfit``
    evaluates the function `(N^2)/2` times to compute the
    coefficients and an additional `N` times to estimate the error.

    **Possible issues**

    One should be careful to use a sufficiently high working
    precision both when calling ``chebyfit`` and when evaluating
    the resulting polynomial, as the polynomial is sometimes
    ill-conditioned. It is for example difficult to reach
    15-digit accuracy when evaluating the polynomial using
    machine precision floats, no matter the theoretical
    accuracy of the polynomial. (The option to return the
    coefficients in Chebyshev form should be made available
    in the future.)

    It is important to note the Chebyshev approximation works
    poorly if `f` is not smooth. A function containing singularities,
    rapid oscillation, etc can be approximated more effectively by
    multiplying it by a weight function that cancels out the
    nonsmooth features, or by dividing the interval into several
    segments.
    """
    a, b = ctx._as_points(interval)
    orig = ctx.prec
    try:
        ctx.prec = orig + int(N**0.5) + 20
        c = [chebcoeff(ctx,f,a,b,k,N) for k in range(N)]
        d = [ctx.zero] * N
        d[0] = -c[0]/2
        h = ctx.mpf(0.5)
        T = chebT(ctx, ctx.mpf(2)/(b-a), ctx.mpf(-1)*(b+a)/(b-a))
        for (k, Tk) in zip(range(N), T):
            for i in range(len(Tk)):
                d[i] += c[k]*Tk[i]
        d = d[::-1]
        # Estimate maximum error
        err = ctx.zero
        for k in range(N):
            x = ctx.cos(ctx.pi*k/N) * (b-a)*h + (b+a)*h
            err = max(err, abs(f(x) - ctx.polyval(d, x)))
    finally:
        ctx.prec = orig
    if error:
        return d, +err
    else:
        return d

@defun
def fourier(ctx, f, interval, N):
    r"""
    Computes the Fourier series of degree `N` of the given function
    on the interval `[a, b]`. More precisely, :func:`~mpmath.fourier` returns
    two lists `(c, s)` of coefficients (the cosine series and sine
    series, respectively), such that

    .. math ::

        f(x) \sim \sum_{k=0}^N
            c_k \cos(k m x) + s_k \sin(k m x)

    where `m = 2 \pi / (b-a)`.

    Note that many texts define the first coefficient as `2 c_0` instead
    of `c_0`. The easiest way to evaluate the computed series correctly
    is to pass it to :func:`~mpmath.fourierval`.

    **Examples**

    The function `f(x) = x` has a simple Fourier series on the standard
    interval `[-\pi, \pi]`. The cosine coefficients are all zero (because
    the function has odd symmetry), and the sine coefficients are
    rational numbers::

        >>> from mpmath import *
        >>> mp.dps = 15; mp.pretty = True
        >>> c, s = fourier(lambda x: x, [-pi, pi], 5)
        >>> nprint(c)
        [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        >>> nprint(s)
        [0.0, 2.0, -1.0, 0.666667, -0.5, 0.4]

    This computes a Fourier series of a nonsymmetric function on
    a nonstandard interval::

        >>> I = [-1, 1.5]
        >>> f = lambda x: x**2 - 4*x + 1
        >>> cs = fourier(f, I, 4)
        >>> nprint(cs[0])
        [0.583333, 1.12479, -1.27552, 0.904708, -0.441296]
        >>> nprint(cs[1])
        [0.0, -2.6255, 0.580905, 0.219974, -0.540057]

    It is instructive to plot a function along with its truncated
    Fourier series::

        >>> plot([f, lambda x: fourierval(cs, I, x)], I) #doctest: +SKIP

    Fourier series generally converge slowly (and may not converge
    pointwise). For example, if `f(x) = \cosh(x)`, a 10-term Fourier
    series gives an `L^2` error corresponding to 2-digit accuracy::

        >>> I = [-1, 1]
        >>> cs = fourier(cosh, I, 9)
        >>> g = lambda x: (cosh(x) - fourierval(cs, I, x))**2
        >>> nprint(sqrt(quad(g, I)))
        0.00467963

    :func:`~mpmath.fourier` uses numerical quadrature. For nonsmooth functions,
    the accuracy (and speed) can be improved by including all singular
    points in the interval specification::

        >>> nprint(fourier(abs, [-1, 1], 0), 10)
        ([0.5000441648], [0.0])
        >>> nprint(fourier(abs, [-1, 0, 1], 0), 10)
        ([0.5], [0.0])

    """
    interval = ctx._as_points(interval)
    a = interval[0]
    b = interval[-1]
    L = b-a
    cos_series = []
    sin_series = []
    cutoff = ctx.eps*10
    for n in xrange(N+1):
        m = 2*n*ctx.pi/L
        an = 2*ctx.quadgl(lambda t: f(t)*ctx.cos(m*t), interval)/L
        bn = 2*ctx.quadgl(lambda t: f(t)*ctx.sin(m*t), interval)/L
        if n == 0:
            an /= 2
        if abs(an) < cutoff: an = ctx.zero
        if abs(bn) < cutoff: bn = ctx.zero
        cos_series.append(an)
        sin_series.append(bn)
    return cos_series, sin_series

@defun
def fourierval(ctx, series, interval, x):
    """
    Evaluates a Fourier series (in the format computed by
    by :func:`~mpmath.fourier` for the given interval) at the point `x`.

    The series should be a pair `(c, s)` where `c` is the
    cosine series and `s` is the sine series. The two lists
    need not have the same length.
    """
    cs, ss = series
    ab = ctx._as_points(interval)
    a = interval[0]
    b = interval[-1]
    m = 2*ctx.pi/(ab[-1]-ab[0])
    s = ctx.zero
    s += ctx.fsum(cs[n]*ctx.cos(m*n*x) for n in xrange(len(cs)) if cs[n])
    s += ctx.fsum(ss[n]*ctx.sin(m*n*x) for n in xrange(len(ss)) if ss[n])
    return s

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\chart\data_source.py ===
"""
Collection of utility primitives for charts.
"""

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Bool,
    Typed,
    Alias,
    String,
    Integer,
    Sequence,
)
from openpyxl.descriptors.excel import ExtensionList
from openpyxl.descriptors.nested import (
    NestedString,
    NestedText,
    NestedInteger,
)


class NumFmt(Serialisable):

    formatCode = String()
    sourceLinked = Bool()

    def __init__(self,
                 formatCode=None,
                 sourceLinked=False
                ):
        self.formatCode = formatCode
        self.sourceLinked = sourceLinked


class NumberValueDescriptor(NestedText):
    """
    Data should be numerical but isn't always :-/
    """

    allow_none = True

    def __set__(self, instance, value):
        if value == "#N/A":
            self.expected_type = str
        else:
            self.expected_type = float
        super().__set__(instance, value)


class NumVal(Serialisable):

    idx = Integer()
    formatCode = NestedText(allow_none=True, expected_type=str)
    v = NumberValueDescriptor()

    def __init__(self,
                 idx=None,
                 formatCode=None,
                 v=None,
                ):
        self.idx = idx
        self.formatCode = formatCode
        self.v = v


class NumData(Serialisable):

    formatCode = NestedText(expected_type=str, allow_none=True)
    ptCount = NestedInteger(allow_none=True)
    pt = Sequence(expected_type=NumVal)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('formatCode', 'ptCount', 'pt')

    def __init__(self,
                 formatCode=None,
                 ptCount=None,
                 pt=(),
                 extLst=None,
                ):
        self.formatCode = formatCode
        self.ptCount = ptCount
        self.pt = pt


class NumRef(Serialisable):

    f = NestedText(expected_type=str)
    ref = Alias('f')
    numCache = Typed(expected_type=NumData, allow_none=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('f', 'numCache')

    def __init__(self,
                 f=None,
                 numCache=None,
                 extLst=None,
                ):
        self.f = f
        self.numCache = numCache


class StrVal(Serialisable):

    tagname = "strVal"

    idx = Integer()
    v = NestedText(expected_type=str)

    def __init__(self,
                 idx=0,
                 v=None,
                ):
        self.idx = idx
        self.v = v


class StrData(Serialisable):

    tagname = "strData"

    ptCount = NestedInteger(allow_none=True)
    pt = Sequence(expected_type=StrVal)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('ptCount', 'pt')

    def __init__(self,
                 ptCount=None,
                 pt=(),
                 extLst=None,
                ):
        self.ptCount = ptCount
        self.pt = pt


class StrRef(Serialisable):

    tagname = "strRef"

    f = NestedText(expected_type=str, allow_none=True)
    strCache = Typed(expected_type=StrData, allow_none=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('f', 'strCache')

    def __init__(self,
                 f=None,
                 strCache=None,
                 extLst=None,
                ):
        self.f = f
        self.strCache = strCache


class NumDataSource(Serialisable):

    numRef = Typed(expected_type=NumRef, allow_none=True)
    numLit = Typed(expected_type=NumData, allow_none=True)


    def __init__(self,
                 numRef=None,
                 numLit=None,
                 ):
        self.numRef = numRef
        self.numLit = numLit


class Level(Serialisable):

    tagname = "lvl"

    pt = Sequence(expected_type=StrVal)

    __elements__ = ('pt',)

    def __init__(self,
                 pt=(),
                ):
        self.pt = pt


class MultiLevelStrData(Serialisable):

    tagname = "multiLvlStrData"

    ptCount = Integer(allow_none=True)
    lvl = Sequence(expected_type=Level)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('ptCount', 'lvl',)

    def __init__(self,
                 ptCount=None,
                 lvl=(),
                 extLst=None,
                ):
        self.ptCount = ptCount
        self.lvl = lvl


class MultiLevelStrRef(Serialisable):

    tagname = "multiLvlStrRef"

    f = NestedText(expected_type=str)
    multiLvlStrCache = Typed(expected_type=MultiLevelStrData, allow_none=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('multiLvlStrCache', 'f')

    def __init__(self,
                 f=None,
                 multiLvlStrCache=None,
                 extLst=None,
                ):
        self.f = f
        self.multiLvlStrCache = multiLvlStrCache


class AxDataSource(Serialisable):

    tagname = "cat"

    numRef = Typed(expected_type=NumRef, allow_none=True)
    numLit = Typed(expected_type=NumData, allow_none=True)
    strRef = Typed(expected_type=StrRef, allow_none=True)
    strLit = Typed(expected_type=StrData, allow_none=True)
    multiLvlStrRef = Typed(expected_type=MultiLevelStrRef, allow_none=True)

    def __init__(self,
                 numRef=None,
                 numLit=None,
                 strRef=None,
                 strLit=None,
                 multiLvlStrRef=None,
                 ):
        if not any([numLit, numRef, strRef, strLit, multiLvlStrRef]):
            raise TypeError("A data source must be provided")
        self.numRef = numRef
        self.numLit = numLit
        self.strRef = strRef
        self.strLit = strLit
        self.multiLvlStrRef = multiLvlStrRef

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\event\legacy.py ===
# event/legacy.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

"""Routines to handle adaption of legacy call signatures,
generation of deprecation notes and docstrings.

"""
from __future__ import annotations

import typing
from typing import Any
from typing import Callable
from typing import List
from typing import Optional
from typing import Tuple
from typing import Type

from .registry import _ET
from .registry import _ListenerFnType
from .. import util
from ..util.compat import FullArgSpec

if typing.TYPE_CHECKING:
    from .attr import _ClsLevelDispatch
    from .base import _HasEventsDispatch


_LegacySignatureType = Tuple[str, List[str], Optional[Callable[..., Any]]]


def _legacy_signature(
    since: str,
    argnames: List[str],
    converter: Optional[Callable[..., Any]] = None,
) -> Callable[[Callable[..., Any]], Callable[..., Any]]:
    """legacy sig decorator


    :param since: string version for deprecation warning
    :param argnames: list of strings, which is *all* arguments that the legacy
     version accepted, including arguments that are still there
    :param converter: lambda that will accept tuple of this full arg signature
     and return tuple of new arg signature.

    """

    def leg(fn: Callable[..., Any]) -> Callable[..., Any]:
        if not hasattr(fn, "_legacy_signatures"):
            fn._legacy_signatures = []  # type: ignore[attr-defined]
        fn._legacy_signatures.append((since, argnames, converter))  # type: ignore[attr-defined] # noqa: E501
        return fn

    return leg


def _wrap_fn_for_legacy(
    dispatch_collection: _ClsLevelDispatch[_ET],
    fn: _ListenerFnType,
    argspec: FullArgSpec,
) -> _ListenerFnType:
    for since, argnames, conv in dispatch_collection.legacy_signatures:
        if argnames[-1] == "**kw":
            has_kw = True
            argnames = argnames[0:-1]
        else:
            has_kw = False

        if len(argnames) == len(argspec.args) and has_kw is bool(
            argspec.varkw
        ):
            formatted_def = "def %s(%s%s)" % (
                dispatch_collection.name,
                ", ".join(dispatch_collection.arg_names),
                ", **kw" if has_kw else "",
            )
            warning_txt = (
                'The argument signature for the "%s.%s" event listener '
                "has changed as of version %s, and conversion for "
                "the old argument signature will be removed in a "
                'future release.  The new signature is "%s"'
                % (
                    dispatch_collection.clsname,
                    dispatch_collection.name,
                    since,
                    formatted_def,
                )
            )

            if conv is not None:
                assert not has_kw

                def wrap_leg(*args: Any, **kw: Any) -> Any:
                    util.warn_deprecated(warning_txt, version=since)
                    assert conv is not None
                    return fn(*conv(*args))

            else:

                def wrap_leg(*args: Any, **kw: Any) -> Any:
                    util.warn_deprecated(warning_txt, version=since)
                    argdict = dict(zip(dispatch_collection.arg_names, args))
                    args_from_dict = [argdict[name] for name in argnames]
                    if has_kw:
                        return fn(*args_from_dict, **kw)
                    else:
                        return fn(*args_from_dict)

            return wrap_leg
    else:
        return fn


def _indent(text: str, indent: str) -> str:
    return "\n".join(indent + line for line in text.split("\n"))


def _standard_listen_example(
    dispatch_collection: _ClsLevelDispatch[_ET],
    sample_target: Any,
    fn: _ListenerFnType,
) -> str:
    example_kw_arg = _indent(
        "\n".join(
            "%(arg)s = kw['%(arg)s']" % {"arg": arg}
            for arg in dispatch_collection.arg_names[0:2]
        ),
        "    ",
    )
    if dispatch_collection.legacy_signatures:
        current_since = max(
            since
            for since, args, conv in dispatch_collection.legacy_signatures
        )
    else:
        current_since = None
    text = (
        "from sqlalchemy import event\n\n\n"
        "@event.listens_for(%(sample_target)s, '%(event_name)s')\n"
        "def receive_%(event_name)s("
        "%(named_event_arguments)s%(has_kw_arguments)s):\n"
        "    \"listen for the '%(event_name)s' event\"\n"
        "\n    # ... (event handling logic) ...\n"
    )

    text %= {
        "current_since": (
            " (arguments as of %s)" % current_since if current_since else ""
        ),
        "event_name": fn.__name__,
        "has_kw_arguments": ", **kw" if dispatch_collection.has_kw else "",
        "named_event_arguments": ", ".join(dispatch_collection.arg_names),
        "example_kw_arg": example_kw_arg,
        "sample_target": sample_target,
    }
    return text


def _legacy_listen_examples(
    dispatch_collection: _ClsLevelDispatch[_ET],
    sample_target: str,
    fn: _ListenerFnType,
) -> str:
    text = ""
    for since, args, conv in dispatch_collection.legacy_signatures:
        text += (
            "\n# DEPRECATED calling style (pre-%(since)s, "
            "will be removed in a future release)\n"
            "@event.listens_for(%(sample_target)s, '%(event_name)s')\n"
            "def receive_%(event_name)s("
            "%(named_event_arguments)s%(has_kw_arguments)s):\n"
            "    \"listen for the '%(event_name)s' event\"\n"
            "\n    # ... (event handling logic) ...\n"
            % {
                "since": since,
                "event_name": fn.__name__,
                "has_kw_arguments": (
                    " **kw" if dispatch_collection.has_kw else ""
                ),
                "named_event_arguments": ", ".join(args),
                "sample_target": sample_target,
            }
        )
    return text


def _version_signature_changes(
    parent_dispatch_cls: Type[_HasEventsDispatch[_ET]],
    dispatch_collection: _ClsLevelDispatch[_ET],
) -> str:
    since, args, conv = dispatch_collection.legacy_signatures[0]
    return (
        "\n.. versionchanged:: %(since)s\n"
        "    The :meth:`.%(clsname)s.%(event_name)s` event now accepts the \n"
        "    arguments %(named_event_arguments)s%(has_kw_arguments)s.\n"
        "    Support for listener functions which accept the previous \n"
        '    argument signature(s) listed above as "deprecated" will be \n'
        "    removed in a future release."
        % {
            "since": since,
            "clsname": parent_dispatch_cls.__name__,
            "event_name": dispatch_collection.name,
            "named_event_arguments": ", ".join(
                ":paramref:`.%(clsname)s.%(event_name)s.%(param_name)s`"
                % {
                    "clsname": parent_dispatch_cls.__name__,
                    "event_name": dispatch_collection.name,
                    "param_name": param_name,
                }
                for param_name in dispatch_collection.arg_names
            ),
            "has_kw_arguments": ", **kw" if dispatch_collection.has_kw else "",
        }
    )


def _augment_fn_docs(
    dispatch_collection: _ClsLevelDispatch[_ET],
    parent_dispatch_cls: Type[_HasEventsDispatch[_ET]],
    fn: _ListenerFnType,
) -> str:
    header = (
        ".. container:: event_signatures\n\n"
        "     Example argument forms::\n"
        "\n"
    )

    sample_target = getattr(parent_dispatch_cls, "_target_class_doc", "obj")
    text = header + _indent(
        _standard_listen_example(dispatch_collection, sample_target, fn),
        " " * 8,
    )
    if dispatch_collection.legacy_signatures:
        text += _indent(
            _legacy_listen_examples(dispatch_collection, sample_target, fn),
            " " * 8,
        )

        text += _version_signature_changes(
            parent_dispatch_cls, dispatch_collection
        )

    return util.inject_docstring_text(fn.__doc__, text, 1)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\numberfields\basis.py ===
"""Computing integral bases for number fields. """

from sympy.polys.polytools import Poly
from sympy.polys.domains.algebraicfield import AlgebraicField
from sympy.polys.domains.integerring import ZZ
from sympy.polys.domains.rationalfield import QQ
from sympy.utilities.decorator import public
from .modules import ModuleEndomorphism, ModuleHomomorphism, PowerBasis
from .utilities import extract_fundamental_discriminant


def _apply_Dedekind_criterion(T, p):
    r"""
    Apply the "Dedekind criterion" to test whether the order needs to be
    enlarged relative to a given prime *p*.
    """
    x = T.gen
    T_bar = Poly(T, modulus=p)
    lc, fl = T_bar.factor_list()
    assert lc == 1
    g_bar = Poly(1, x, modulus=p)
    for ti_bar, _ in fl:
        g_bar *= ti_bar
    h_bar = T_bar // g_bar
    g = Poly(g_bar, domain=ZZ)
    h = Poly(h_bar, domain=ZZ)
    f = (g * h - T) // p
    f_bar = Poly(f, modulus=p)
    Z_bar = f_bar
    for b in [g_bar, h_bar]:
        Z_bar = Z_bar.gcd(b)
    U_bar = T_bar // Z_bar
    m = Z_bar.degree()
    return U_bar, m


def nilradical_mod_p(H, p, q=None):
    r"""
    Compute the nilradical mod *p* for a given order *H*, and prime *p*.

    Explanation
    ===========

    This is the ideal $I$ in $H/pH$ consisting of all elements some positive
    power of which is zero in this quotient ring, i.e. is a multiple of *p*.

    Parameters
    ==========

    H : :py:class:`~.Submodule`
        The given order.
    p : int
        The rational prime.
    q : int, optional
        If known, the smallest power of *p* that is $>=$ the dimension of *H*.
        If not provided, we compute it here.

    Returns
    =======

    :py:class:`~.Module` representing the nilradical mod *p* in *H*.

    References
    ==========

    .. [1] Cohen, H. *A Course in Computational Algebraic Number Theory*.
    (See Lemma 6.1.6.)

    """
    n = H.n
    if q is None:
        q = p
        while q < n:
            q *= p
    phi = ModuleEndomorphism(H, lambda x: x**q)
    return phi.kernel(modulus=p)


def _second_enlargement(H, p, q):
    r"""
    Perform the second enlargement in the Round Two algorithm.
    """
    Ip = nilradical_mod_p(H, p, q=q)
    B = H.parent.submodule_from_matrix(H.matrix * Ip.matrix, denom=H.denom)
    C = B + p*H
    E = C.endomorphism_ring()
    phi = ModuleHomomorphism(H, E, lambda x: E.inner_endomorphism(x))
    gamma = phi.kernel(modulus=p)
    G = H.parent.submodule_from_matrix(H.matrix * gamma.matrix, denom=H.denom * p)
    H1 = G + H
    return H1, Ip


@public
def round_two(T, radicals=None):
    r"""
    Zassenhaus's "Round 2" algorithm.

    Explanation
    ===========

    Carry out Zassenhaus's "Round 2" algorithm on an irreducible polynomial
    *T* over :ref:`ZZ` or :ref:`QQ`. This computes an integral basis and the
    discriminant for the field $K = \mathbb{Q}[x]/(T(x))$.

    Alternatively, you may pass an :py:class:`~.AlgebraicField` instance, in
    place of the polynomial *T*, in which case the algorithm is applied to the
    minimal polynomial for the field's primitive element.

    Ordinarily this function need not be called directly, as one can instead
    access the :py:meth:`~.AlgebraicField.maximal_order`,
    :py:meth:`~.AlgebraicField.integral_basis`, and
    :py:meth:`~.AlgebraicField.discriminant` methods of an
    :py:class:`~.AlgebraicField`.

    Examples
    ========

    Working through an AlgebraicField:

    >>> from sympy import Poly, QQ
    >>> from sympy.abc import x
    >>> T = Poly(x ** 3 + x ** 2 - 2 * x + 8)
    >>> K = QQ.alg_field_from_poly(T, "theta")
    >>> print(K.maximal_order())
    Submodule[[2, 0, 0], [0, 2, 0], [0, 1, 1]]/2
    >>> print(K.discriminant())
    -503
    >>> print(K.integral_basis(fmt='sympy'))
    [1, theta, theta/2 + theta**2/2]

    Calling directly:

    >>> from sympy import Poly
    >>> from sympy.abc import x
    >>> from sympy.polys.numberfields.basis import round_two
    >>> T = Poly(x ** 3 + x ** 2 - 2 * x + 8)
    >>> print(round_two(T))
    (Submodule[[2, 0, 0], [0, 2, 0], [0, 1, 1]]/2, -503)

    The nilradicals mod $p$ that are sometimes computed during the Round Two
    algorithm may be useful in further calculations. Pass a dictionary under
    `radicals` to receive these:

    >>> T = Poly(x**3 + 3*x**2 + 5)
    >>> rad = {}
    >>> ZK, dK = round_two(T, radicals=rad)
    >>> print(rad)
    {3: Submodule[[-1, 1, 0], [-1, 0, 1]]}

    Parameters
    ==========

    T : :py:class:`~.Poly`, :py:class:`~.AlgebraicField`
        Either (1) the irreducible polynomial over :ref:`ZZ` or :ref:`QQ`
        defining the number field, or (2) an :py:class:`~.AlgebraicField`
        representing the number field itself.

    radicals : dict, optional
        This is a way for any $p$-radicals (if computed) to be returned by
        reference. If desired, pass an empty dictionary. If the algorithm
        reaches the point where it computes the nilradical mod $p$ of the ring
        of integers $Z_K$, then an $\mathbb{F}_p$-basis for this ideal will be
        stored in this dictionary under the key ``p``. This can be useful for
        other algorithms, such as prime decomposition.

    Returns
    =======

    Pair ``(ZK, dK)``, where:

        ``ZK`` is a :py:class:`~sympy.polys.numberfields.modules.Submodule`
        representing the maximal order.

        ``dK`` is the discriminant of the field $K = \mathbb{Q}[x]/(T(x))$.

    See Also
    ========

    .AlgebraicField.maximal_order
    .AlgebraicField.integral_basis
    .AlgebraicField.discriminant

    References
    ==========

    .. [1] Cohen, H. *A Course in Computational Algebraic Number Theory.*

    """
    K = None
    if isinstance(T, AlgebraicField):
        K, T = T, T.ext.minpoly_of_element()
    if (   not T.is_univariate
        or not T.is_irreducible
        or T.domain not in [ZZ, QQ]):
        raise ValueError('Round 2 requires an irreducible univariate polynomial over ZZ or QQ.')
    T, _ = T.make_monic_over_integers_by_scaling_roots()
    n = T.degree()
    D = T.discriminant()
    D_modulus = ZZ.from_sympy(abs(D))
    # D must be 0 or 1 mod 4 (see Cohen Sec 4.4), which ensures we can write
    # it in the form D = D_0 * F**2, where D_0 is 1 or a fundamental discriminant.
    _, F = extract_fundamental_discriminant(D)
    Ztheta = PowerBasis(K or T)
    H = Ztheta.whole_submodule()
    nilrad = None
    while F:
        # Next prime:
        p, e = F.popitem()
        U_bar, m = _apply_Dedekind_criterion(T, p)
        if m == 0:
            continue
        # For a given prime p, the first enlargement of the order spanned by
        # the current basis can be done in a simple way:
        U = Ztheta.element_from_poly(Poly(U_bar, domain=ZZ))
        # TODO:
        #  Theory says only first m columns of the U//p*H term below are needed.
        #  Could be slightly more efficient to use only those. Maybe `Submodule`
        #  class should support a slice operator?
        H = H.add(U // p * H, hnf_modulus=D_modulus)
        if e <= m:
            continue
        # A second, and possibly more, enlargements for p will be needed.
        # These enlargements require a more involved procedure.
        q = p
        while q < n:
            q *= p
        H1, nilrad = _second_enlargement(H, p, q)
        while H1 != H:
            H = H1
            H1, nilrad = _second_enlargement(H, p, q)
    # Note: We do not store all nilradicals mod p, only the very last. This is
    # because, unless computed against the entire integral basis, it might not
    # be accurate. (In other words, if H was not already equal to ZK when we
    # passed it to `_second_enlargement`, then we can't trust the nilradical
    # so computed.) Example: if T(x) = x ** 3 + 15 * x ** 2 - 9 * x + 13, then
    # F is divisible by 2, 3, and 7, and the nilradical mod 2 as computed above
    # will not be accurate for the full, maximal order ZK.
    if nilrad is not None and isinstance(radicals, dict):
        radicals[p] = nilrad
    ZK = H
    # Pre-set expensive boolean properties which we already know to be true:
    ZK._starts_with_unity = True
    ZK._is_sq_maxrank_HNF = True
    dK = (D * ZK.matrix.det() ** 2) // ZK.denom ** (2 * n)
    return ZK, dK

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\sets\conditionset.py ===
from sympy.core.singleton import S
from sympy.core.basic import Basic
from sympy.core.containers import Tuple
from sympy.core.function import Lambda, BadSignatureError
from sympy.core.logic import fuzzy_bool
from sympy.core.relational import Eq
from sympy.core.symbol import Dummy
from sympy.core.sympify import _sympify
from sympy.logic.boolalg import And, as_Boolean
from sympy.utilities.iterables import sift, flatten, has_dups
from sympy.utilities.exceptions import sympy_deprecation_warning
from .contains import Contains
from .sets import Set, Union, FiniteSet, SetKind


adummy = Dummy('conditionset')


class ConditionSet(Set):
    r"""
    Set of elements which satisfies a given condition.

    .. math:: \{x \mid \textrm{condition}(x) = \texttt{True}, x \in S\}

    Examples
    ========

    >>> from sympy import Symbol, S, ConditionSet, pi, Eq, sin, Interval
    >>> from sympy.abc import x, y, z

    >>> sin_sols = ConditionSet(x, Eq(sin(x), 0), Interval(0, 2*pi))
    >>> 2*pi in sin_sols
    True
    >>> pi/2 in sin_sols
    False
    >>> 3*pi in sin_sols
    False
    >>> 5 in ConditionSet(x, x**2 > 4, S.Reals)
    True

    If the value is not in the base set, the result is false:

    >>> 5 in ConditionSet(x, x**2 > 4, Interval(2, 4))
    False

    Notes
    =====

    Symbols with assumptions should be avoided or else the
    condition may evaluate without consideration of the set:

    >>> n = Symbol('n', negative=True)
    >>> cond = (n > 0); cond
    False
    >>> ConditionSet(n, cond, S.Integers)
    EmptySet

    Only free symbols can be changed by using `subs`:

    >>> c = ConditionSet(x, x < 1, {x, z})
    >>> c.subs(x, y)
    ConditionSet(x, x < 1, {y, z})

    To check if ``pi`` is in ``c`` use:

    >>> pi in c
    False

    If no base set is specified, the universal set is implied:

    >>> ConditionSet(x, x < 1).base_set
    UniversalSet

    Only symbols or symbol-like expressions can be used:

    >>> ConditionSet(x + 1, x + 1 < 1, S.Integers)
    Traceback (most recent call last):
    ...
    ValueError: non-symbol dummy not recognized in condition

    When the base set is a ConditionSet, the symbols will be
    unified if possible with preference for the outermost symbols:

    >>> ConditionSet(x, x < y, ConditionSet(z, z + y < 2, S.Integers))
    ConditionSet(x, (x < y) & (x + y < 2), Integers)

    """
    def __new__(cls, sym, condition, base_set=S.UniversalSet):
        sym = _sympify(sym)
        flat = flatten([sym])
        if has_dups(flat):
            raise BadSignatureError("Duplicate symbols detected")
        base_set = _sympify(base_set)
        if not isinstance(base_set, Set):
            raise TypeError(
                'base set should be a Set object, not %s' % base_set)
        condition = _sympify(condition)

        if isinstance(condition, FiniteSet):
            condition_orig = condition
            temp = (Eq(lhs, 0) for lhs in condition)
            condition = And(*temp)
            sympy_deprecation_warning(
                f"""
Using a set for the condition in ConditionSet is deprecated. Use a boolean
instead.

In this case, replace

    {condition_orig}

with

    {condition}
""",
                deprecated_since_version='1.5',
                active_deprecations_target="deprecated-conditionset-set",
                )

        condition = as_Boolean(condition)

        if condition is S.true:
            return base_set

        if condition is S.false:
            return S.EmptySet

        if base_set is S.EmptySet:
            return S.EmptySet

        # no simple answers, so now check syms
        for i in flat:
            if not getattr(i, '_diff_wrt', False):
                raise ValueError('`%s` is not symbol-like' % i)

        if base_set.contains(sym) is S.false:
            raise TypeError('sym `%s` is not in base_set `%s`' % (sym, base_set))

        know = None
        if isinstance(base_set, FiniteSet):
            sifted = sift(
                base_set, lambda _: fuzzy_bool(condition.subs(sym, _)))
            if sifted[None]:
                know = FiniteSet(*sifted[True])
                base_set = FiniteSet(*sifted[None])
            else:
                return FiniteSet(*sifted[True])

        if isinstance(base_set, cls):
            s, c, b = base_set.args
            def sig(s):
                return cls(s, Eq(adummy, 0)).as_dummy().sym
            sa, sb = map(sig, (sym, s))
            if sa != sb:
                raise BadSignatureError('sym does not match sym of base set')
            reps = dict(zip(flatten([sym]), flatten([s])))
            if s == sym:
                condition = And(condition, c)
                base_set = b
            elif not c.free_symbols & sym.free_symbols:
                reps = {v: k for k, v in reps.items()}
                condition = And(condition, c.xreplace(reps))
                base_set = b
            elif not condition.free_symbols & s.free_symbols:
                sym = sym.xreplace(reps)
                condition = And(condition.xreplace(reps), c)
                base_set = b

        # flatten ConditionSet(Contains(ConditionSet())) expressions
        if isinstance(condition, Contains) and (sym == condition.args[0]):
            if isinstance(condition.args[1], Set):
                return condition.args[1].intersect(base_set)

        rv = Basic.__new__(cls, sym, condition, base_set)
        return rv if know is None else Union(know, rv)

    sym = property(lambda self: self.args[0])
    condition = property(lambda self: self.args[1])
    base_set = property(lambda self: self.args[2])

    @property
    def free_symbols(self):
        cond_syms = self.condition.free_symbols - self.sym.free_symbols
        return cond_syms | self.base_set.free_symbols

    @property
    def bound_symbols(self):
        return flatten([self.sym])

    def _contains(self, other):
        def ok_sig(a, b):
            tuples = [isinstance(i, Tuple) for i in (a, b)]
            c = tuples.count(True)
            if c == 1:
                return False
            if c == 0:
                return True
            return len(a) == len(b) and all(
                ok_sig(i, j) for i, j in zip(a, b))
        if not ok_sig(self.sym, other):
            return S.false

        # try doing base_cond first and return
        # False immediately if it is False
        base_cond = Contains(other, self.base_set)
        if base_cond is S.false:
            return S.false

        # Substitute other into condition. This could raise e.g. for
        # ConditionSet(x, 1/x >= 0, Reals).contains(0)
        lamda = Lambda((self.sym,), self.condition)
        try:
            lambda_cond = lamda(other)
        except TypeError:
            return None
        else:
            return And(base_cond, lambda_cond)

    def as_relational(self, other):
        f = Lambda(self.sym, self.condition)
        if isinstance(self.sym, Tuple):
            f = f(*other)
        else:
            f = f(other)
        return And(f, self.base_set.contains(other))

    def _eval_subs(self, old, new):
        sym, cond, base = self.args
        dsym = sym.subs(old, adummy)
        insym = dsym.has(adummy)
        # prioritize changing a symbol in the base
        newbase = base.subs(old, new)
        if newbase != base:
            if not insym:
                cond = cond.subs(old, new)
            return self.func(sym, cond, newbase)
        if insym:
            pass  # no change of bound symbols via subs
        elif getattr(new, '_diff_wrt', False):
            cond = cond.subs(old, new)
        else:
            pass  # let error about the symbol raise from __new__
        return self.func(sym, cond, base)

    def _kind(self):
        return SetKind(self.sym.kind)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\check_type_completeness.py ===
#!/usr/bin/env python3
"""This is a file that wraps calls to `pyright --verifytypes`, achieving two things:
1. give an error if docstrings are missing.
    pyright will give a number of missing docstrings, and error messages, but not exit with a non-zero value.
2. filter out specific errors we don't care about.
    this is largely due to 1, but also because Trio does some very complex stuff and --verifytypes has few to no ways of ignoring specific errors.

If this check is giving you false alarms, you can ignore them by adding logic to `has_docstring_at_runtime`, in the main loop in `check_type`, or by updating the json file.
"""
from __future__ import annotations

# this file is not run as part of the tests, instead it's run standalone from check.sh
import argparse
import json
import subprocess
import sys
from pathlib import Path

import trio
import trio.testing

# not needed if everything is working, but if somebody does something to generate
# tons of errors, we can be nice and stop them from getting 3*tons of output
printed_diagnostics: set[str] = set()


# TODO: consider checking manually without `--ignoreexternal`, and/or
# removing it from the below call later on.
def run_pyright(platform: str) -> subprocess.CompletedProcess[bytes]:
    return subprocess.run(
        [
            "pyright",
            # Specify a platform and version to keep imported modules consistent.
            f"--pythonplatform={platform}",
            "--pythonversion=3.9",
            "--verifytypes=trio",
            "--outputjson",
            "--ignoreexternal",
        ],
        capture_output=True,
    )


def has_docstring_at_runtime(name: str) -> bool:
    """Pyright gives us an object identifier of xx.yy.zz
    This function tries to decompose that into its constituent parts, such that we
    can resolve it, in order to check whether it has a `__doc__` at runtime and
    verifytypes misses it because we're doing overly fancy stuff.
    """
    # This assert is solely for stopping isort from removing our imports of trio & trio.testing
    # It could also be done with isort:skip, but that'd also disable import sorting and the like.
    assert trio.testing is not None

    # figure out what part of the name is the module, so we can "import" it
    name_parts = name.split(".")
    assert name_parts[0] == "trio"
    if name_parts[1] == "tests":
        return True

    # traverse down the remaining identifiers with getattr
    obj = trio
    try:
        for obj_name in name_parts[1:]:
            obj = getattr(obj, obj_name)
    except AttributeError as exc:
        # asynciowrapper does funky getattr stuff
        if "AsyncIOWrapper" in str(exc) or name in (
            # Symbols not existing on all platforms, so we can't dynamically inspect them.
            # Manually confirmed to have docstrings but pyright doesn't see them due to
            # export shenanigans. TODO: actually manually confirm that.
            # In theory we could verify these at runtime, probably by running the script separately
            # on separate platforms. It might also be a decent idea to work the other way around,
            # a la test_static_tool_sees_class_members
            # darwin
            "trio.lowlevel.current_kqueue",
            "trio.lowlevel.monitor_kevent",
            "trio.lowlevel.wait_kevent",
            "trio._core._io_kqueue._KqueueStatistics",
            # windows
            "trio._socket.SocketType.share",
            "trio._core._io_windows._WindowsStatistics",
            "trio._core._windows_cffi.Handle",
            "trio.lowlevel.current_iocp",
            "trio.lowlevel.monitor_completion_key",
            "trio.lowlevel.readinto_overlapped",
            "trio.lowlevel.register_with_iocp",
            "trio.lowlevel.wait_overlapped",
            "trio.lowlevel.write_overlapped",
            "trio.lowlevel.WaitForSingleObject",
            "trio.socket.fromshare",
            # linux
            # this test will fail on linux, but I don't develop on linux. So the next
            # person to do so is very welcome to open a pull request and populate with
            # objects
            # TODO: these are erroring on all platforms, why?
            "trio._highlevel_generic.StapledStream.send_stream",
            "trio._highlevel_generic.StapledStream.receive_stream",
            "trio._ssl.SSLStream.transport_stream",
            "trio._file_io._HasFileNo",
            "trio._file_io._HasFileNo.fileno",
        ):
            return True

        else:
            print(
                f"Pyright sees {name} at runtime, but unable to getattr({obj.__name__}, {obj_name}).",
                file=sys.stderr,
            )
            return False
    return bool(obj.__doc__)


def check_type(
    platform: str,
    full_diagnostics_file: Path | None,
    expected_errors: list[object],
) -> list[object]:
    # convince isort we use the trio import
    assert trio is not None

    # run pyright, load output into json
    res = run_pyright(platform)
    current_result = json.loads(res.stdout)

    if res.stderr:
        print(res.stderr, file=sys.stderr)

    if full_diagnostics_file:
        with open(full_diagnostics_file, "a") as f:
            json.dump(current_result, f, sort_keys=True, indent=4)

    errors = []

    for symbol in current_result["typeCompleteness"]["symbols"]:
        diagnostics = symbol["diagnostics"]
        name = symbol["name"]
        for diagnostic in diagnostics:
            message = diagnostic["message"]
            if name in (
                "trio._path.PosixPath",
                "trio._path.WindowsPath",
            ) and message.startswith("Type of base class "):
                continue

            if name.startswith("trio._path.Path"):
                if message.startswith("No docstring found for"):
                    continue
                if message.startswith(
                    "Type is missing type annotation and could be inferred differently by type checkers",
                ):
                    continue

            # ignore errors about missing docstrings if they're available at runtime
            if message.startswith("No docstring found for"):
                if has_docstring_at_runtime(symbol["name"]):
                    continue
            else:
                # Missing docstring messages include the name of the object.
                # Other errors don't, so we add it.
                message = f"{name}: {message}"
            if message not in expected_errors and message not in printed_diagnostics:
                print(f"new error: {message}", file=sys.stderr)
            errors.append(message)
            printed_diagnostics.add(message)

        continue

    return errors


def main(args: argparse.Namespace) -> int:
    if args.full_diagnostics_file:
        full_diagnostics_file = Path(args.full_diagnostics_file)
        full_diagnostics_file.write_text("")
    else:
        full_diagnostics_file = None

    errors_by_platform_file = Path(__file__).parent / "_check_type_completeness.json"
    if errors_by_platform_file.exists():
        with open(errors_by_platform_file) as f:
            errors_by_platform = json.load(f)
    else:
        errors_by_platform = {"Linux": [], "Windows": [], "Darwin": [], "all": []}

    changed = False
    for platform in "Linux", "Windows", "Darwin":
        platform_errors = errors_by_platform[platform] + errors_by_platform["all"]
        print("*" * 20, f"\nChecking {platform}...")
        errors = check_type(platform, full_diagnostics_file, platform_errors)

        new_errors = [e for e in errors if e not in platform_errors]
        missing_errors = [e for e in platform_errors if e not in errors]

        if new_errors:
            print(
                f"New errors introduced in `pyright --verifytypes`. Fix them, or ignore them by modifying {errors_by_platform_file}, either manually or with '--overwrite-file'.",
                file=sys.stderr,
            )
            changed = True
        if missing_errors:
            print(
                f"Congratulations, you have resolved existing errors! Please remove them from {errors_by_platform_file}, either manually or with '--overwrite-file'.",
                file=sys.stderr,
            )
            changed = True
            print(missing_errors, file=sys.stderr)

        errors_by_platform[platform] = errors
    print("*" * 20)

    # cut down the size of the json file by a lot, and make it easier to parse for
    # humans, by moving errors that appear on all platforms to a separate category
    errors_by_platform["all"] = []
    for e in errors_by_platform["Linux"].copy():
        if e in errors_by_platform["Darwin"] and e in errors_by_platform["Windows"]:
            for platform in "Linux", "Windows", "Darwin":
                errors_by_platform[platform].remove(e)
            errors_by_platform["all"].append(e)

    if changed and args.overwrite_file:
        with open(errors_by_platform_file, "w") as f:
            json.dump(errors_by_platform, f, indent=4, sort_keys=True)
            # newline at end of file
            f.write("\n")

    # True -> 1 -> non-zero exit value -> error
    return changed


parser = argparse.ArgumentParser()
parser.add_argument(
    "--overwrite-file",
    action="store_true",
    default=False,
    help="Use this flag to overwrite the current stored results. Either in CI together with a diff check, or to avoid having to manually correct it.",
)
parser.add_argument(
    "--full-diagnostics-file",
    type=Path,
    default=None,
    help="Use this for debugging, it will dump the output of all three pyright runs by platform into this file.",
)
args = parser.parse_args()

assert __name__ == "__main__", "This script should be run standalone"
sys.exit(main(args))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\google\protobuf\internal\wire_format.py ===
# Protocol Buffers - Google's data interchange format
# Copyright 2008 Google Inc.  All rights reserved.
#
# Use of this source code is governed by a BSD-style
# license that can be found in the LICENSE file or at
# https://developers.google.com/open-source/licenses/bsd

"""Constants and static functions to support protocol buffer wire format."""

__author__ = 'robinson@google.com (Will Robinson)'

import struct
from google.protobuf import descriptor
from google.protobuf import message


TAG_TYPE_BITS = 3  # Number of bits used to hold type info in a proto tag.
TAG_TYPE_MASK = (1 << TAG_TYPE_BITS) - 1  # 0x7

# These numbers identify the wire type of a protocol buffer value.
# We use the least-significant TAG_TYPE_BITS bits of the varint-encoded
# tag-and-type to store one of these WIRETYPE_* constants.
# These values must match WireType enum in //google/protobuf/wire_format.h.
WIRETYPE_VARINT = 0
WIRETYPE_FIXED64 = 1
WIRETYPE_LENGTH_DELIMITED = 2
WIRETYPE_START_GROUP = 3
WIRETYPE_END_GROUP = 4
WIRETYPE_FIXED32 = 5
_WIRETYPE_MAX = 5


# Bounds for various integer types.
INT32_MAX = int((1 << 31) - 1)
INT32_MIN = int(-(1 << 31))
UINT32_MAX = (1 << 32) - 1

INT64_MAX = (1 << 63) - 1
INT64_MIN = -(1 << 63)
UINT64_MAX = (1 << 64) - 1

# "struct" format strings that will encode/decode the specified formats.
FORMAT_UINT32_LITTLE_ENDIAN = '<I'
FORMAT_UINT64_LITTLE_ENDIAN = '<Q'
FORMAT_FLOAT_LITTLE_ENDIAN = '<f'
FORMAT_DOUBLE_LITTLE_ENDIAN = '<d'


# We'll have to provide alternate implementations of AppendLittleEndian*() on
# any architectures where these checks fail.
if struct.calcsize(FORMAT_UINT32_LITTLE_ENDIAN) != 4:
  raise AssertionError('Format "I" is not a 32-bit number.')
if struct.calcsize(FORMAT_UINT64_LITTLE_ENDIAN) != 8:
  raise AssertionError('Format "Q" is not a 64-bit number.')


def PackTag(field_number, wire_type):
  """Returns an unsigned 32-bit integer that encodes the field number and
  wire type information in standard protocol message wire format.

  Args:
    field_number: Expected to be an integer in the range [1, 1 << 29)
    wire_type: One of the WIRETYPE_* constants.
  """
  if not 0 <= wire_type <= _WIRETYPE_MAX:
    raise message.EncodeError('Unknown wire type: %d' % wire_type)
  return (field_number << TAG_TYPE_BITS) | wire_type


def UnpackTag(tag):
  """The inverse of PackTag().  Given an unsigned 32-bit number,
  returns a (field_number, wire_type) tuple.
  """
  return (tag >> TAG_TYPE_BITS), (tag & TAG_TYPE_MASK)


def ZigZagEncode(value):
  """ZigZag Transform:  Encodes signed integers so that they can be
  effectively used with varint encoding.  See wire_format.h for
  more details.
  """
  if value >= 0:
    return value << 1
  return (value << 1) ^ (~0)


def ZigZagDecode(value):
  """Inverse of ZigZagEncode()."""
  if not value & 0x1:
    return value >> 1
  return (value >> 1) ^ (~0)



# The *ByteSize() functions below return the number of bytes required to
# serialize "field number + type" information and then serialize the value.


def Int32ByteSize(field_number, int32):
  return Int64ByteSize(field_number, int32)


def Int32ByteSizeNoTag(int32):
  return _VarUInt64ByteSizeNoTag(0xffffffffffffffff & int32)


def Int64ByteSize(field_number, int64):
  # Have to convert to uint before calling UInt64ByteSize().
  return UInt64ByteSize(field_number, 0xffffffffffffffff & int64)


def UInt32ByteSize(field_number, uint32):
  return UInt64ByteSize(field_number, uint32)


def UInt64ByteSize(field_number, uint64):
  return TagByteSize(field_number) + _VarUInt64ByteSizeNoTag(uint64)


def SInt32ByteSize(field_number, int32):
  return UInt32ByteSize(field_number, ZigZagEncode(int32))


def SInt64ByteSize(field_number, int64):
  return UInt64ByteSize(field_number, ZigZagEncode(int64))


def Fixed32ByteSize(field_number, fixed32):
  return TagByteSize(field_number) + 4


def Fixed64ByteSize(field_number, fixed64):
  return TagByteSize(field_number) + 8


def SFixed32ByteSize(field_number, sfixed32):
  return TagByteSize(field_number) + 4


def SFixed64ByteSize(field_number, sfixed64):
  return TagByteSize(field_number) + 8


def FloatByteSize(field_number, flt):
  return TagByteSize(field_number) + 4


def DoubleByteSize(field_number, double):
  return TagByteSize(field_number) + 8


def BoolByteSize(field_number, b):
  return TagByteSize(field_number) + 1


def EnumByteSize(field_number, enum):
  return UInt32ByteSize(field_number, enum)


def StringByteSize(field_number, string):
  return BytesByteSize(field_number, string.encode('utf-8'))


def BytesByteSize(field_number, b):
  return (TagByteSize(field_number)
          + _VarUInt64ByteSizeNoTag(len(b))
          + len(b))


def GroupByteSize(field_number, message):
  return (2 * TagByteSize(field_number)  # START and END group.
          + message.ByteSize())


def MessageByteSize(field_number, message):
  return (TagByteSize(field_number)
          + _VarUInt64ByteSizeNoTag(message.ByteSize())
          + message.ByteSize())


def MessageSetItemByteSize(field_number, msg):
  # First compute the sizes of the tags.
  # There are 2 tags for the beginning and ending of the repeated group, that
  # is field number 1, one with field number 2 (type_id) and one with field
  # number 3 (message).
  total_size = (2 * TagByteSize(1) + TagByteSize(2) + TagByteSize(3))

  # Add the number of bytes for type_id.
  total_size += _VarUInt64ByteSizeNoTag(field_number)

  message_size = msg.ByteSize()

  # The number of bytes for encoding the length of the message.
  total_size += _VarUInt64ByteSizeNoTag(message_size)

  # The size of the message.
  total_size += message_size
  return total_size


def TagByteSize(field_number):
  """Returns the bytes required to serialize a tag with this field number."""
  # Just pass in type 0, since the type won't affect the tag+type size.
  return _VarUInt64ByteSizeNoTag(PackTag(field_number, 0))


# Private helper function for the *ByteSize() functions above.

def _VarUInt64ByteSizeNoTag(uint64):
  """Returns the number of bytes required to serialize a single varint
  using boundary value comparisons. (unrolled loop optimization -WPierce)
  uint64 must be unsigned.
  """
  if uint64 <= 0x7f: return 1
  if uint64 <= 0x3fff: return 2
  if uint64 <= 0x1fffff: return 3
  if uint64 <= 0xfffffff: return 4
  if uint64 <= 0x7ffffffff: return 5
  if uint64 <= 0x3ffffffffff: return 6
  if uint64 <= 0x1ffffffffffff: return 7
  if uint64 <= 0xffffffffffffff: return 8
  if uint64 <= 0x7fffffffffffffff: return 9
  if uint64 > UINT64_MAX:
    raise message.EncodeError('Value out of range: %d' % uint64)
  return 10


NON_PACKABLE_TYPES = (
  descriptor.FieldDescriptor.TYPE_STRING,
  descriptor.FieldDescriptor.TYPE_GROUP,
  descriptor.FieldDescriptor.TYPE_MESSAGE,
  descriptor.FieldDescriptor.TYPE_BYTES
)


def IsTypePackable(field_type):
  """Return true iff packable = true is valid for fields of this type.

  Args:
    field_type: a FieldDescriptor::Type value.

  Returns:
    True iff fields of this type are packable.
  """
  return field_type not in NON_PACKABLE_TYPES

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\_numba\kernels\var_.py ===
"""
Numba 1D var kernels that can be shared by
* Dataframe / Series
* groupby
* rolling / expanding

Mirrors pandas/_libs/window/aggregation.pyx
"""
from __future__ import annotations

from typing import TYPE_CHECKING

import numba
import numpy as np

if TYPE_CHECKING:
    from pandas._typing import npt

from pandas.core._numba.kernels.shared import is_monotonic_increasing


@numba.jit(nopython=True, nogil=True, parallel=False)
def add_var(
    val: float,
    nobs: int,
    mean_x: float,
    ssqdm_x: float,
    compensation: float,
    num_consecutive_same_value: int,
    prev_value: float,
) -> tuple[int, float, float, float, int, float]:
    if not np.isnan(val):
        if val == prev_value:
            num_consecutive_same_value += 1
        else:
            num_consecutive_same_value = 1
        prev_value = val

        nobs += 1
        prev_mean = mean_x - compensation
        y = val - compensation
        t = y - mean_x
        compensation = t + mean_x - y
        delta = t
        if nobs:
            mean_x += delta / nobs
        else:
            mean_x = 0
        ssqdm_x += (val - prev_mean) * (val - mean_x)
    return nobs, mean_x, ssqdm_x, compensation, num_consecutive_same_value, prev_value


@numba.jit(nopython=True, nogil=True, parallel=False)
def remove_var(
    val: float, nobs: int, mean_x: float, ssqdm_x: float, compensation: float
) -> tuple[int, float, float, float]:
    if not np.isnan(val):
        nobs -= 1
        if nobs:
            prev_mean = mean_x - compensation
            y = val - compensation
            t = y - mean_x
            compensation = t + mean_x - y
            delta = t
            mean_x -= delta / nobs
            ssqdm_x -= (val - prev_mean) * (val - mean_x)
        else:
            mean_x = 0
            ssqdm_x = 0
    return nobs, mean_x, ssqdm_x, compensation


@numba.jit(nopython=True, nogil=True, parallel=False)
def sliding_var(
    values: np.ndarray,
    result_dtype: np.dtype,
    start: np.ndarray,
    end: np.ndarray,
    min_periods: int,
    ddof: int = 1,
) -> tuple[np.ndarray, list[int]]:
    N = len(start)
    nobs = 0
    mean_x = 0.0
    ssqdm_x = 0.0
    compensation_add = 0.0
    compensation_remove = 0.0

    min_periods = max(min_periods, 1)
    is_monotonic_increasing_bounds = is_monotonic_increasing(
        start
    ) and is_monotonic_increasing(end)

    output = np.empty(N, dtype=result_dtype)

    for i in range(N):
        s = start[i]
        e = end[i]
        if i == 0 or not is_monotonic_increasing_bounds:
            prev_value = values[s]
            num_consecutive_same_value = 0

            for j in range(s, e):
                val = values[j]
                (
                    nobs,
                    mean_x,
                    ssqdm_x,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,
                ) = add_var(
                    val,
                    nobs,
                    mean_x,
                    ssqdm_x,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,
                )
        else:
            for j in range(start[i - 1], s):
                val = values[j]
                nobs, mean_x, ssqdm_x, compensation_remove = remove_var(
                    val, nobs, mean_x, ssqdm_x, compensation_remove
                )

            for j in range(end[i - 1], e):
                val = values[j]
                (
                    nobs,
                    mean_x,
                    ssqdm_x,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,
                ) = add_var(
                    val,
                    nobs,
                    mean_x,
                    ssqdm_x,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,
                )

        if nobs >= min_periods and nobs > ddof:
            if nobs == 1 or num_consecutive_same_value >= nobs:
                result = 0.0
            else:
                result = ssqdm_x / (nobs - ddof)
        else:
            result = np.nan

        output[i] = result

        if not is_monotonic_increasing_bounds:
            nobs = 0
            mean_x = 0.0
            ssqdm_x = 0.0
            compensation_remove = 0.0

    # na_position is empty list since float64 can already hold nans
    # Do list comprehension, since numba cannot figure out that na_pos is
    # empty list of ints on its own
    na_pos = [0 for i in range(0)]
    return output, na_pos


@numba.jit(nopython=True, nogil=True, parallel=False)
def grouped_var(
    values: np.ndarray,
    result_dtype: np.dtype,
    labels: npt.NDArray[np.intp],
    ngroups: int,
    min_periods: int,
    ddof: int = 1,
) -> tuple[np.ndarray, list[int]]:
    N = len(labels)

    nobs_arr = np.zeros(ngroups, dtype=np.int64)
    comp_arr = np.zeros(ngroups, dtype=values.dtype)
    consecutive_counts = np.zeros(ngroups, dtype=np.int64)
    prev_vals = np.zeros(ngroups, dtype=values.dtype)
    output = np.zeros(ngroups, dtype=result_dtype)
    means = np.zeros(ngroups, dtype=result_dtype)

    for i in range(N):
        lab = labels[i]
        val = values[i]

        if lab < 0:
            continue

        mean_x = means[lab]
        ssqdm_x = output[lab]
        nobs = nobs_arr[lab]
        compensation_add = comp_arr[lab]
        num_consecutive_same_value = consecutive_counts[lab]
        prev_value = prev_vals[lab]

        (
            nobs,
            mean_x,
            ssqdm_x,
            compensation_add,
            num_consecutive_same_value,
            prev_value,
        ) = add_var(
            val,
            nobs,
            mean_x,
            ssqdm_x,
            compensation_add,
            num_consecutive_same_value,
            prev_value,
        )

        output[lab] = ssqdm_x
        means[lab] = mean_x
        consecutive_counts[lab] = num_consecutive_same_value
        prev_vals[lab] = prev_value
        comp_arr[lab] = compensation_add
        nobs_arr[lab] = nobs

    # Post-processing, replace vars that don't satisfy min_periods
    for lab in range(ngroups):
        nobs = nobs_arr[lab]
        num_consecutive_same_value = consecutive_counts[lab]
        ssqdm_x = output[lab]
        if nobs >= min_periods and nobs > ddof:
            if nobs == 1 or num_consecutive_same_value >= nobs:
                result = 0.0
            else:
                result = ssqdm_x / (nobs - ddof)
        else:
            result = np.nan
        output[lab] = result

    # Second pass to get the std.dev
    # na_position is empty list since float64 can already hold nans
    # Do list comprehension, since numba cannot figure out that na_pos is
    # empty list of ints on its own
    na_pos = [0 for i in range(0)]
    return output, na_pos

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\resolution\resolvelib\requirements.py ===
from typing import Any, Optional

from pip._vendor.packaging.specifiers import SpecifierSet
from pip._vendor.packaging.utils import NormalizedName, canonicalize_name

from pip._internal.req.constructors import install_req_drop_extras
from pip._internal.req.req_install import InstallRequirement

from .base import Candidate, CandidateLookup, Requirement, format_name


class ExplicitRequirement(Requirement):
    def __init__(self, candidate: Candidate) -> None:
        self.candidate = candidate

    def __str__(self) -> str:
        return str(self.candidate)

    def __repr__(self) -> str:
        return f"{self.__class__.__name__}({self.candidate!r})"

    def __hash__(self) -> int:
        return hash(self.candidate)

    def __eq__(self, other: Any) -> bool:
        if not isinstance(other, ExplicitRequirement):
            return False
        return self.candidate == other.candidate

    @property
    def project_name(self) -> NormalizedName:
        # No need to canonicalize - the candidate did this
        return self.candidate.project_name

    @property
    def name(self) -> str:
        # No need to canonicalize - the candidate did this
        return self.candidate.name

    def format_for_error(self) -> str:
        return self.candidate.format_for_error()

    def get_candidate_lookup(self) -> CandidateLookup:
        return self.candidate, None

    def is_satisfied_by(self, candidate: Candidate) -> bool:
        return candidate == self.candidate


class SpecifierRequirement(Requirement):
    def __init__(self, ireq: InstallRequirement) -> None:
        assert ireq.link is None, "This is a link, not a specifier"
        self._ireq = ireq
        self._equal_cache: Optional[str] = None
        self._hash: Optional[int] = None
        self._extras = frozenset(canonicalize_name(e) for e in self._ireq.extras)

    @property
    def _equal(self) -> str:
        if self._equal_cache is not None:
            return self._equal_cache

        self._equal_cache = str(self._ireq)
        return self._equal_cache

    def __str__(self) -> str:
        return str(self._ireq.req)

    def __repr__(self) -> str:
        return f"{self.__class__.__name__}({str(self._ireq.req)!r})"

    def __eq__(self, other: object) -> bool:
        if not isinstance(other, SpecifierRequirement):
            return NotImplemented
        return self._equal == other._equal

    def __hash__(self) -> int:
        if self._hash is not None:
            return self._hash

        self._hash = hash(self._equal)
        return self._hash

    @property
    def project_name(self) -> NormalizedName:
        assert self._ireq.req, "Specifier-backed ireq is always PEP 508"
        return canonicalize_name(self._ireq.req.name)

    @property
    def name(self) -> str:
        return format_name(self.project_name, self._extras)

    def format_for_error(self) -> str:
        # Convert comma-separated specifiers into "A, B, ..., F and G"
        # This makes the specifier a bit more "human readable", without
        # risking a change in meaning. (Hopefully! Not all edge cases have
        # been checked)
        parts = [s.strip() for s in str(self).split(",")]
        if len(parts) == 0:
            return ""
        elif len(parts) == 1:
            return parts[0]

        return ", ".join(parts[:-1]) + " and " + parts[-1]

    def get_candidate_lookup(self) -> CandidateLookup:
        return None, self._ireq

    def is_satisfied_by(self, candidate: Candidate) -> bool:
        assert candidate.name == self.name, (
            f"Internal issue: Candidate is not for this requirement "
            f"{candidate.name} vs {self.name}"
        )
        # We can safely always allow prereleases here since PackageFinder
        # already implements the prerelease logic, and would have filtered out
        # prerelease candidates if the user does not expect them.
        assert self._ireq.req, "Specifier-backed ireq is always PEP 508"
        spec = self._ireq.req.specifier
        return spec.contains(candidate.version, prereleases=True)


class SpecifierWithoutExtrasRequirement(SpecifierRequirement):
    """
    Requirement backed by an install requirement on a base package.
    Trims extras from its install requirement if there are any.
    """

    def __init__(self, ireq: InstallRequirement) -> None:
        assert ireq.link is None, "This is a link, not a specifier"
        self._ireq = install_req_drop_extras(ireq)
        self._equal_cache: Optional[str] = None
        self._hash: Optional[int] = None
        self._extras = frozenset(canonicalize_name(e) for e in self._ireq.extras)

    @property
    def _equal(self) -> str:
        if self._equal_cache is not None:
            return self._equal_cache

        self._equal_cache = str(self._ireq)
        return self._equal_cache

    def __eq__(self, other: object) -> bool:
        if not isinstance(other, SpecifierWithoutExtrasRequirement):
            return NotImplemented
        return self._equal == other._equal

    def __hash__(self) -> int:
        if self._hash is not None:
            return self._hash

        self._hash = hash(self._equal)
        return self._hash


class RequiresPythonRequirement(Requirement):
    """A requirement representing Requires-Python metadata."""

    def __init__(self, specifier: SpecifierSet, match: Candidate) -> None:
        self.specifier = specifier
        self._specifier_string = str(specifier)  # for faster __eq__
        self._hash: Optional[int] = None
        self._candidate = match

    def __str__(self) -> str:
        return f"Python {self.specifier}"

    def __repr__(self) -> str:
        return f"{self.__class__.__name__}({str(self.specifier)!r})"

    def __hash__(self) -> int:
        if self._hash is not None:
            return self._hash

        self._hash = hash((self._specifier_string, self._candidate))
        return self._hash

    def __eq__(self, other: Any) -> bool:
        if not isinstance(other, RequiresPythonRequirement):
            return False
        return (
            self._specifier_string == other._specifier_string
            and self._candidate == other._candidate
        )

    @property
    def project_name(self) -> NormalizedName:
        return self._candidate.project_name

    @property
    def name(self) -> str:
        return self._candidate.name

    def format_for_error(self) -> str:
        return str(self)

    def get_candidate_lookup(self) -> CandidateLookup:
        if self.specifier.contains(self._candidate.version, prereleases=True):
            return self._candidate, None
        return None, None

    def is_satisfied_by(self, candidate: Candidate) -> bool:
        assert candidate.name == self._candidate.name, "Not Python candidate"
        # We can safely always allow prereleases here since PackageFinder
        # already implements the prerelease logic, and would have filtered out
        # prerelease candidates if the user does not expect them.
        return self.specifier.contains(candidate.version, prereleases=True)


class UnsatisfiableRequirement(Requirement):
    """A requirement that cannot be satisfied."""

    def __init__(self, name: NormalizedName) -> None:
        self._name = name

    def __str__(self) -> str:
        return f"{self._name} (unavailable)"

    def __repr__(self) -> str:
        return f"{self.__class__.__name__}({str(self._name)!r})"

    def __eq__(self, other: object) -> bool:
        if not isinstance(other, UnsatisfiableRequirement):
            return NotImplemented
        return self._name == other._name

    def __hash__(self) -> int:
        return hash(self._name)

    @property
    def project_name(self) -> NormalizedName:
        return self._name

    @property
    def name(self) -> str:
        return self._name

    def format_for_error(self) -> str:
        return str(self)

    def get_candidate_lookup(self) -> CandidateLookup:
        return None, None

    def is_satisfied_by(self, candidate: Candidate) -> bool:
        return False

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\utils\subprocess.py ===
import logging
import os
import shlex
import subprocess
from typing import Any, Callable, Iterable, List, Literal, Mapping, Optional, Union

from pip._vendor.rich.markup import escape

from pip._internal.cli.spinners import SpinnerInterface, open_spinner
from pip._internal.exceptions import InstallationSubprocessError
from pip._internal.utils.logging import VERBOSE, subprocess_logger
from pip._internal.utils.misc import HiddenText

CommandArgs = List[Union[str, HiddenText]]


def make_command(*args: Union[str, HiddenText, CommandArgs]) -> CommandArgs:
    """
    Create a CommandArgs object.
    """
    command_args: CommandArgs = []
    for arg in args:
        # Check for list instead of CommandArgs since CommandArgs is
        # only known during type-checking.
        if isinstance(arg, list):
            command_args.extend(arg)
        else:
            # Otherwise, arg is str or HiddenText.
            command_args.append(arg)

    return command_args


def format_command_args(args: Union[List[str], CommandArgs]) -> str:
    """
    Format command arguments for display.
    """
    # For HiddenText arguments, display the redacted form by calling str().
    # Also, we don't apply str() to arguments that aren't HiddenText since
    # this can trigger a UnicodeDecodeError in Python 2 if the argument
    # has type unicode and includes a non-ascii character.  (The type
    # checker doesn't ensure the annotations are correct in all cases.)
    return " ".join(
        shlex.quote(str(arg)) if isinstance(arg, HiddenText) else shlex.quote(arg)
        for arg in args
    )


def reveal_command_args(args: Union[List[str], CommandArgs]) -> List[str]:
    """
    Return the arguments in their raw, unredacted form.
    """
    return [arg.secret if isinstance(arg, HiddenText) else arg for arg in args]


def call_subprocess(
    cmd: Union[List[str], CommandArgs],
    show_stdout: bool = False,
    cwd: Optional[str] = None,
    on_returncode: 'Literal["raise", "warn", "ignore"]' = "raise",
    extra_ok_returncodes: Optional[Iterable[int]] = None,
    extra_environ: Optional[Mapping[str, Any]] = None,
    unset_environ: Optional[Iterable[str]] = None,
    spinner: Optional[SpinnerInterface] = None,
    log_failed_cmd: Optional[bool] = True,
    stdout_only: Optional[bool] = False,
    *,
    command_desc: str,
) -> str:
    """
    Args:
      show_stdout: if true, use INFO to log the subprocess's stderr and
        stdout streams.  Otherwise, use DEBUG.  Defaults to False.
      extra_ok_returncodes: an iterable of integer return codes that are
        acceptable, in addition to 0. Defaults to None, which means [].
      unset_environ: an iterable of environment variable names to unset
        prior to calling subprocess.Popen().
      log_failed_cmd: if false, failed commands are not logged, only raised.
      stdout_only: if true, return only stdout, else return both. When true,
        logging of both stdout and stderr occurs when the subprocess has
        terminated, else logging occurs as subprocess output is produced.
    """
    if extra_ok_returncodes is None:
        extra_ok_returncodes = []
    if unset_environ is None:
        unset_environ = []
    # Most places in pip use show_stdout=False. What this means is--
    #
    # - We connect the child's output (combined stderr and stdout) to a
    #   single pipe, which we read.
    # - We log this output to stderr at DEBUG level as it is received.
    # - If DEBUG logging isn't enabled (e.g. if --verbose logging wasn't
    #   requested), then we show a spinner so the user can still see the
    #   subprocess is in progress.
    # - If the subprocess exits with an error, we log the output to stderr
    #   at ERROR level if it hasn't already been displayed to the console
    #   (e.g. if --verbose logging wasn't enabled).  This way we don't log
    #   the output to the console twice.
    #
    # If show_stdout=True, then the above is still done, but with DEBUG
    # replaced by INFO.
    if show_stdout:
        # Then log the subprocess output at INFO level.
        log_subprocess: Callable[..., None] = subprocess_logger.info
        used_level = logging.INFO
    else:
        # Then log the subprocess output using VERBOSE.  This also ensures
        # it will be logged to the log file (aka user_log), if enabled.
        log_subprocess = subprocess_logger.verbose
        used_level = VERBOSE

    # Whether the subprocess will be visible in the console.
    showing_subprocess = subprocess_logger.getEffectiveLevel() <= used_level

    # Only use the spinner if we're not showing the subprocess output
    # and we have a spinner.
    use_spinner = not showing_subprocess and spinner is not None

    log_subprocess("Running command %s", command_desc)
    env = os.environ.copy()
    if extra_environ:
        env.update(extra_environ)
    for name in unset_environ:
        env.pop(name, None)
    try:
        proc = subprocess.Popen(
            # Convert HiddenText objects to the underlying str.
            reveal_command_args(cmd),
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT if not stdout_only else subprocess.PIPE,
            cwd=cwd,
            env=env,
            errors="backslashreplace",
        )
    except Exception as exc:
        if log_failed_cmd:
            subprocess_logger.critical(
                "Error %s while executing command %s",
                exc,
                command_desc,
            )
        raise
    all_output = []
    if not stdout_only:
        assert proc.stdout
        assert proc.stdin
        proc.stdin.close()
        # In this mode, stdout and stderr are in the same pipe.
        while True:
            line: str = proc.stdout.readline()
            if not line:
                break
            line = line.rstrip()
            all_output.append(line + "\n")

            # Show the line immediately.
            log_subprocess(line)
            # Update the spinner.
            if use_spinner:
                assert spinner
                spinner.spin()
        try:
            proc.wait()
        finally:
            if proc.stdout:
                proc.stdout.close()
        output = "".join(all_output)
    else:
        # In this mode, stdout and stderr are in different pipes.
        # We must use communicate() which is the only safe way to read both.
        out, err = proc.communicate()
        # log line by line to preserve pip log indenting
        for out_line in out.splitlines():
            log_subprocess(out_line)
        all_output.append(out)
        for err_line in err.splitlines():
            log_subprocess(err_line)
        all_output.append(err)
        output = out

    proc_had_error = proc.returncode and proc.returncode not in extra_ok_returncodes
    if use_spinner:
        assert spinner
        if proc_had_error:
            spinner.finish("error")
        else:
            spinner.finish("done")
    if proc_had_error:
        if on_returncode == "raise":
            error = InstallationSubprocessError(
                command_description=command_desc,
                exit_code=proc.returncode,
                output_lines=all_output if not showing_subprocess else None,
            )
            if log_failed_cmd:
                subprocess_logger.error("%s", error, extra={"rich": True})
                subprocess_logger.verbose(
                    "[bold magenta]full command[/]: [blue]%s[/]",
                    escape(format_command_args(cmd)),
                    extra={"markup": True},
                )
                subprocess_logger.verbose(
                    "[bold magenta]cwd[/]: %s",
                    escape(cwd or "[inherit]"),
                    extra={"markup": True},
                )

            raise error
        elif on_returncode == "warn":
            subprocess_logger.warning(
                'Command "%s" had error code %s in %s',
                command_desc,
                proc.returncode,
                cwd,
            )
        elif on_returncode == "ignore":
            pass
        else:
            raise ValueError(f"Invalid value: on_returncode={on_returncode!r}")
    return output


def runner_with_spinner_message(message: str) -> Callable[..., None]:
    """Provide a subprocess_runner that shows a spinner message.

    Intended for use with for BuildBackendHookCaller. Thus, the runner has
    an API that matches what's expected by BuildBackendHookCaller.subprocess_runner.
    """

    def runner(
        cmd: List[str],
        cwd: Optional[str] = None,
        extra_environ: Optional[Mapping[str, Any]] = None,
    ) -> None:
        with open_spinner(message) as spinner:
            call_subprocess(
                cmd,
                command_desc=message,
                cwd=cwd,
                extra_environ=extra_environ,
                spinner=spinner,
            )

    return runner

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\service.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.
import errno
import logging
import os
import subprocess
from abc import ABC, abstractmethod
from io import IOBase
from platform import system
from subprocess import PIPE
from time import sleep
from typing import IO, Any, List, Mapping, Optional, Union, cast
from urllib import request
from urllib.error import URLError

from selenium.common.exceptions import WebDriverException
from selenium.types import SubprocessStdAlias
from selenium.webdriver.common import utils

logger = logging.getLogger(__name__)


class Service(ABC):
    """The abstract base class for all service objects.  Services typically
    launch a child program in a new process as an interim process to
    communicate with a browser.

    :param executable: install path of the executable.
    :param port: Port for the service to run on, defaults to 0 where the operating system will decide.
    :param log_output: (Optional) int representation of STDOUT/DEVNULL, any IO instance or String path to file.
    :param env: (Optional) Mapping of environment variables for the new process, defaults to `os.environ`.
    :param driver_path_env_key: (Optional) Environment variable to use to get the path to the driver executable.
    """

    def __init__(
        self,
        executable_path: Optional[str] = None,
        port: int = 0,
        log_output: Optional[SubprocessStdAlias] = None,
        env: Optional[Mapping[Any, Any]] = None,
        driver_path_env_key: Optional[str] = None,
        **kwargs,
    ) -> None:
        if isinstance(log_output, str):
            self.log_output = cast(IOBase, open(log_output, "a+", encoding="utf-8"))
        elif log_output == subprocess.STDOUT:
            self.log_output = cast(Optional[Union[int, IOBase]], None)
        elif log_output is None or log_output == subprocess.DEVNULL:
            self.log_output = cast(Optional[Union[int, IOBase]], subprocess.DEVNULL)
        else:
            self.log_output = log_output

        self.port = port or utils.free_port()
        # Default value for every python subprocess: subprocess.Popen(..., creationflags=0)
        self.popen_kw = kwargs.pop("popen_kw", {})
        self.creation_flags = self.popen_kw.pop("creation_flags", 0)
        self.env = env or os.environ
        self.DRIVER_PATH_ENV_KEY = driver_path_env_key
        self._path = self.env_path() or executable_path

    @property
    def service_url(self) -> str:
        """Gets the url of the Service."""
        return f"http://{utils.join_host_port('localhost', self.port)}"

    @abstractmethod
    def command_line_args(self) -> List[str]:
        """A List of program arguments (excluding the executable)."""
        raise NotImplementedError("This method needs to be implemented in a sub class")

    @property
    def path(self) -> str:
        return self._path or ""

    @path.setter
    def path(self, value: str) -> None:
        self._path = str(value)

    def start(self) -> None:
        """Starts the Service.

        :Exceptions:
         - WebDriverException : Raised either when it can't start the service
           or when it can't connect to the service
        """
        if self._path is None:
            raise WebDriverException("Service path cannot be None.")
        self._start_process(self._path)

        count = 0
        while True:
            self.assert_process_still_running()
            if self.is_connectable():
                break
            # sleep increasing: 0.01, 0.06, 0.11, 0.16, 0.21, 0.26, 0.31, 0.36, 0.41, 0.46, 0.5
            sleep(min(0.01 + 0.05 * count, 0.5))
            count += 1
            if count == 70:
                raise WebDriverException(f"Can not connect to the Service {self._path}")

    def assert_process_still_running(self) -> None:
        """Check if the underlying process is still running."""
        return_code = self.process.poll()
        if return_code:
            raise WebDriverException(f"Service {self._path} unexpectedly exited. Status code was: {return_code}")

    def is_connectable(self) -> bool:
        """Establishes a socket connection to determine if the service running
        on the port is accessible."""
        return utils.is_connectable(self.port)

    def send_remote_shutdown_command(self) -> None:
        """Dispatch an HTTP request to the shutdown endpoint for the service in
        an attempt to stop it."""
        try:
            request.urlopen(f"{self.service_url}/shutdown")
        except URLError:
            return

        for _ in range(30):
            if not self.is_connectable():
                break
            sleep(1)

    def stop(self) -> None:
        """Stops the service."""

        if self.log_output not in {PIPE, subprocess.DEVNULL}:
            if isinstance(self.log_output, IOBase):
                self.log_output.close()
            elif isinstance(self.log_output, int):
                os.close(self.log_output)

        if self.process is not None and self.process.poll() is None:
            try:
                self.send_remote_shutdown_command()
            except TypeError:
                pass
            finally:
                self._terminate_process()

    def _terminate_process(self) -> None:
        """Terminate the child process.

        On POSIX this attempts a graceful SIGTERM followed by a SIGKILL,
        on a Windows OS kill is an alias to terminate.  Terminating does
        not raise itself if something has gone wrong but (currently)
        silently ignores errors here.
        """
        try:
            stdin, stdout, stderr = (
                self.process.stdin,
                self.process.stdout,
                self.process.stderr,
            )
            for stream in stdin, stdout, stderr:
                try:
                    stream.close()  # type: ignore
                except AttributeError:
                    pass
            self.process.terminate()
            try:
                self.process.wait(60)
            except subprocess.TimeoutExpired:
                logger.error(
                    "Service process refused to terminate gracefully with SIGTERM, escalating to SIGKILL.",
                    exc_info=True,
                )
                self.process.kill()
        except OSError:
            logger.error("Error terminating service process.", exc_info=True)

    def __del__(self) -> None:
        # `subprocess.Popen` doesn't send signal on `__del__`;
        # so we attempt to close the launched process when `__del__`
        # is triggered.
        # do not use globals here; interpreter shutdown may have already cleaned them up
        # and they would be `None`. This goes for anything this method is referencing internally.
        try:
            self.stop()
        except Exception:
            pass

    def _start_process(self, path: str) -> None:
        """Creates a subprocess by executing the command provided.

        :param cmd: full command to execute
        """
        cmd = [path]
        cmd.extend(self.command_line_args())
        close_file_descriptors = self.popen_kw.pop("close_fds", system() != "Windows")
        try:
            start_info = None
            if system() == "Windows":
                start_info = subprocess.STARTUPINFO()  # type: ignore[attr-defined]
                start_info.dwFlags = subprocess.CREATE_NEW_CONSOLE | subprocess.STARTF_USESHOWWINDOW  # type: ignore[attr-defined]
                start_info.wShowWindow = subprocess.SW_HIDE  # type: ignore[attr-defined]

            self.process = subprocess.Popen(
                cmd,
                env=self.env,
                close_fds=close_file_descriptors,
                stdout=cast(Optional[Union[int, IO[Any]]], self.log_output),
                stderr=cast(Optional[Union[int, IO[Any]]], self.log_output),
                stdin=PIPE,
                creationflags=self.creation_flags,
                startupinfo=start_info,
                **self.popen_kw,
            )
            logger.debug(
                "Started executable: `%s` in a child process with pid: %s using %s to output %s",
                self._path,
                self.process.pid,
                self.creation_flags,
                self.log_output,
            )
        except TypeError:
            raise
        except OSError as err:
            if err.errno == errno.EACCES:
                if self._path is None:
                    raise WebDriverException("Service path cannot be None.")
                raise WebDriverException(
                    f"'{os.path.basename(self._path)}' executable may have wrong permissions."
                ) from err
            raise

    def env_path(self) -> Optional[str]:
        if self.DRIVER_PATH_ENV_KEY:
            return os.getenv(self.DRIVER_PATH_ENV_KEY, None)
        return None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\_numba\kernels\sum_.py ===
"""
Numba 1D sum kernels that can be shared by
* Dataframe / Series
* groupby
* rolling / expanding

Mirrors pandas/_libs/window/aggregation.pyx
"""
from __future__ import annotations

from typing import (
    TYPE_CHECKING,
    Any,
)

import numba
from numba.extending import register_jitable
import numpy as np

if TYPE_CHECKING:
    from pandas._typing import npt

from pandas.core._numba.kernels.shared import is_monotonic_increasing


@numba.jit(nopython=True, nogil=True, parallel=False)
def add_sum(
    val: Any,
    nobs: int,
    sum_x: Any,
    compensation: Any,
    num_consecutive_same_value: int,
    prev_value: Any,
) -> tuple[int, Any, Any, int, Any]:
    if not np.isnan(val):
        nobs += 1
        y = val - compensation
        t = sum_x + y
        compensation = t - sum_x - y
        sum_x = t

        if val == prev_value:
            num_consecutive_same_value += 1
        else:
            num_consecutive_same_value = 1
        prev_value = val

    return nobs, sum_x, compensation, num_consecutive_same_value, prev_value


@numba.jit(nopython=True, nogil=True, parallel=False)
def remove_sum(
    val: Any, nobs: int, sum_x: Any, compensation: Any
) -> tuple[int, Any, Any]:
    if not np.isnan(val):
        nobs -= 1
        y = -val - compensation
        t = sum_x + y
        compensation = t - sum_x - y
        sum_x = t
    return nobs, sum_x, compensation


@numba.jit(nopython=True, nogil=True, parallel=False)
def sliding_sum(
    values: np.ndarray,
    result_dtype: np.dtype,
    start: np.ndarray,
    end: np.ndarray,
    min_periods: int,
) -> tuple[np.ndarray, list[int]]:
    dtype = values.dtype

    na_val: object = np.nan
    if dtype.kind == "i":
        na_val = 0

    N = len(start)
    nobs = 0
    sum_x = 0
    compensation_add = 0
    compensation_remove = 0
    na_pos = []

    is_monotonic_increasing_bounds = is_monotonic_increasing(
        start
    ) and is_monotonic_increasing(end)

    output = np.empty(N, dtype=result_dtype)

    for i in range(N):
        s = start[i]
        e = end[i]
        if i == 0 or not is_monotonic_increasing_bounds:
            prev_value = values[s]
            num_consecutive_same_value = 0

            for j in range(s, e):
                val = values[j]
                (
                    nobs,
                    sum_x,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,
                ) = add_sum(
                    val,
                    nobs,
                    sum_x,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,
                )
        else:
            for j in range(start[i - 1], s):
                val = values[j]
                nobs, sum_x, compensation_remove = remove_sum(
                    val, nobs, sum_x, compensation_remove
                )

            for j in range(end[i - 1], e):
                val = values[j]
                (
                    nobs,
                    sum_x,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,
                ) = add_sum(
                    val,
                    nobs,
                    sum_x,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,
                )

        if nobs == 0 == min_periods:
            result: object = 0
        elif nobs >= min_periods:
            if num_consecutive_same_value >= nobs:
                result = prev_value * nobs
            else:
                result = sum_x
        else:
            result = na_val
            if dtype.kind == "i":
                na_pos.append(i)

        output[i] = result

        if not is_monotonic_increasing_bounds:
            nobs = 0
            sum_x = 0
            compensation_remove = 0

    return output, na_pos


# Mypy/pyright don't like the fact that the decorator is untyped
@register_jitable  # type: ignore[misc]
def grouped_kahan_sum(
    values: np.ndarray,
    result_dtype: np.dtype,
    labels: npt.NDArray[np.intp],
    ngroups: int,
) -> tuple[
    np.ndarray, npt.NDArray[np.int64], np.ndarray, npt.NDArray[np.int64], np.ndarray
]:
    N = len(labels)

    nobs_arr = np.zeros(ngroups, dtype=np.int64)
    comp_arr = np.zeros(ngroups, dtype=values.dtype)
    consecutive_counts = np.zeros(ngroups, dtype=np.int64)
    prev_vals = np.zeros(ngroups, dtype=values.dtype)
    output = np.zeros(ngroups, dtype=result_dtype)

    for i in range(N):
        lab = labels[i]
        val = values[i]

        if lab < 0:
            continue

        sum_x = output[lab]
        nobs = nobs_arr[lab]
        compensation_add = comp_arr[lab]
        num_consecutive_same_value = consecutive_counts[lab]
        prev_value = prev_vals[lab]

        (
            nobs,
            sum_x,
            compensation_add,
            num_consecutive_same_value,
            prev_value,
        ) = add_sum(
            val,
            nobs,
            sum_x,
            compensation_add,
            num_consecutive_same_value,
            prev_value,
        )

        output[lab] = sum_x
        consecutive_counts[lab] = num_consecutive_same_value
        prev_vals[lab] = prev_value
        comp_arr[lab] = compensation_add
        nobs_arr[lab] = nobs
    return output, nobs_arr, comp_arr, consecutive_counts, prev_vals


@numba.jit(nopython=True, nogil=True, parallel=False)
def grouped_sum(
    values: np.ndarray,
    result_dtype: np.dtype,
    labels: npt.NDArray[np.intp],
    ngroups: int,
    min_periods: int,
) -> tuple[np.ndarray, list[int]]:
    na_pos = []

    output, nobs_arr, comp_arr, consecutive_counts, prev_vals = grouped_kahan_sum(
        values, result_dtype, labels, ngroups
    )

    # Post-processing, replace sums that don't satisfy min_periods
    for lab in range(ngroups):
        nobs = nobs_arr[lab]
        num_consecutive_same_value = consecutive_counts[lab]
        prev_value = prev_vals[lab]
        sum_x = output[lab]
        if nobs >= min_periods:
            if num_consecutive_same_value >= nobs:
                result = prev_value * nobs
            else:
                result = sum_x
        else:
            result = sum_x  # Don't change val, will be replaced by nan later
            na_pos.append(lab)
        output[lab] = result

    return output, na_pos

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\websocket\_wsdump.py ===
#!/usr/bin/env python3

"""
wsdump.py
websocket - WebSocket client library for Python

Copyright 2024 engn33r

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

import argparse
import code
import gzip
import ssl
import sys
import threading
import time
import zlib
from urllib.parse import urlparse

import websocket

try:
    import readline
except ImportError:
    pass


def get_encoding() -> str:
    encoding = getattr(sys.stdin, "encoding", "")
    if not encoding:
        return "utf-8"
    else:
        return encoding.lower()


OPCODE_DATA = (websocket.ABNF.OPCODE_TEXT, websocket.ABNF.OPCODE_BINARY)
ENCODING = get_encoding()


class VAction(argparse.Action):
    def __call__(
        self,
        parser: argparse.Namespace,
        args: tuple,
        values: str,
        option_string: str = None,
    ) -> None:
        if values is None:
            values = "1"
        try:
            values = int(values)
        except ValueError:
            values = values.count("v") + 1
        setattr(args, self.dest, values)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="WebSocket Simple Dump Tool")
    parser.add_argument(
        "url", metavar="ws_url", help="websocket url. ex. ws://echo.websocket.events/"
    )
    parser.add_argument("-p", "--proxy", help="proxy url. ex. http://127.0.0.1:8080")
    parser.add_argument(
        "-v",
        "--verbose",
        default=0,
        nargs="?",
        action=VAction,
        dest="verbose",
        help="set verbose mode. If set to 1, show opcode. "
        "If set to 2, enable to trace  websocket module",
    )
    parser.add_argument(
        "-n", "--nocert", action="store_true", help="Ignore invalid SSL cert"
    )
    parser.add_argument("-r", "--raw", action="store_true", help="raw output")
    parser.add_argument("-s", "--subprotocols", nargs="*", help="Set subprotocols")
    parser.add_argument("-o", "--origin", help="Set origin")
    parser.add_argument(
        "--eof-wait",
        default=0,
        type=int,
        help="wait time(second) after 'EOF' received.",
    )
    parser.add_argument("-t", "--text", help="Send initial text")
    parser.add_argument(
        "--timings", action="store_true", help="Print timings in seconds"
    )
    parser.add_argument("--headers", help="Set custom headers. Use ',' as separator")

    return parser.parse_args()


class RawInput:
    def raw_input(self, prompt: str = "") -> str:
        line = input(prompt)

        if ENCODING and ENCODING != "utf-8" and not isinstance(line, str):
            line = line.decode(ENCODING).encode("utf-8")
        elif isinstance(line, str):
            line = line.encode("utf-8")

        return line


class InteractiveConsole(RawInput, code.InteractiveConsole):
    def write(self, data: str) -> None:
        sys.stdout.write("\033[2K\033[E")
        # sys.stdout.write("\n")
        sys.stdout.write("\033[34m< " + data + "\033[39m")
        sys.stdout.write("\n> ")
        sys.stdout.flush()

    def read(self) -> str:
        return self.raw_input("> ")


class NonInteractive(RawInput):
    def write(self, data: str) -> None:
        sys.stdout.write(data)
        sys.stdout.write("\n")
        sys.stdout.flush()

    def read(self) -> str:
        return self.raw_input("")


def main() -> None:
    start_time = time.time()
    args = parse_args()
    if args.verbose > 1:
        websocket.enableTrace(True)
    options = {}
    if args.proxy:
        p = urlparse(args.proxy)
        options["http_proxy_host"] = p.hostname
        options["http_proxy_port"] = p.port
    if args.origin:
        options["origin"] = args.origin
    if args.subprotocols:
        options["subprotocols"] = args.subprotocols
    opts = {}
    if args.nocert:
        opts = {"cert_reqs": ssl.CERT_NONE, "check_hostname": False}
    if args.headers:
        options["header"] = list(map(str.strip, args.headers.split(",")))
    ws = websocket.create_connection(args.url, sslopt=opts, **options)
    if args.raw:
        console = NonInteractive()
    else:
        console = InteractiveConsole()
        print("Press Ctrl+C to quit")

    def recv() -> tuple:
        try:
            frame = ws.recv_frame()
        except websocket.WebSocketException:
            return websocket.ABNF.OPCODE_CLOSE, ""
        if not frame:
            raise websocket.WebSocketException(f"Not a valid frame {frame}")
        elif frame.opcode in OPCODE_DATA:
            return frame.opcode, frame.data
        elif frame.opcode == websocket.ABNF.OPCODE_CLOSE:
            ws.send_close()
            return frame.opcode, ""
        elif frame.opcode == websocket.ABNF.OPCODE_PING:
            ws.pong(frame.data)
            return frame.opcode, frame.data

        return frame.opcode, frame.data

    def recv_ws() -> None:
        while True:
            opcode, data = recv()
            msg = None
            if opcode == websocket.ABNF.OPCODE_TEXT and isinstance(data, bytes):
                data = str(data, "utf-8")
            if (
                isinstance(data, bytes) and len(data) > 2 and data[:2] == b"\037\213"
            ):  # gzip magick
                try:
                    data = "[gzip] " + str(gzip.decompress(data), "utf-8")
                except:
                    pass
            elif isinstance(data, bytes):
                try:
                    data = "[zlib] " + str(
                        zlib.decompress(data, -zlib.MAX_WBITS), "utf-8"
                    )
                except:
                    pass

            if isinstance(data, bytes):
                data = repr(data)

            if args.verbose:
                msg = f"{websocket.ABNF.OPCODE_MAP.get(opcode)}: {data}"
            else:
                msg = data

            if msg is not None:
                if args.timings:
                    console.write(f"{time.time() - start_time}: {msg}")
                else:
                    console.write(msg)

            if opcode == websocket.ABNF.OPCODE_CLOSE:
                break

    thread = threading.Thread(target=recv_ws)
    thread.daemon = True
    thread.start()

    if args.text:
        ws.send(args.text)

    while True:
        try:
            message = console.read()
            ws.send(message)
        except KeyboardInterrupt:
            return
        except EOFError:
            time.sleep(args.eof_wait)
            return


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(e)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\mysql\enumerated.py ===
# dialects/mysql/enumerated.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors


import re

from .types import _StringType
from ... import exc
from ... import sql
from ... import util
from ...sql import sqltypes


class ENUM(sqltypes.NativeForEmulated, sqltypes.Enum, _StringType):
    """MySQL ENUM type."""

    __visit_name__ = "ENUM"

    native_enum = True

    def __init__(self, *enums, **kw):
        """Construct an ENUM.

        E.g.::

          Column("myenum", ENUM("foo", "bar", "baz"))

        :param enums: The range of valid values for this ENUM.  Values in
          enums are not quoted, they will be escaped and surrounded by single
          quotes when generating the schema.  This object may also be a
          PEP-435-compliant enumerated type.

          .. versionadded: 1.1 added support for PEP-435-compliant enumerated
             types.

        :param strict: This flag has no effect.

         .. versionchanged:: The MySQL ENUM type as well as the base Enum
            type now validates all Python data values.

        :param charset: Optional, a column-level character set for this string
          value.  Takes precedence to 'ascii' or 'unicode' short-hand.

        :param collation: Optional, a column-level collation for this string
          value.  Takes precedence to 'binary' short-hand.

        :param ascii: Defaults to False: short-hand for the ``latin1``
          character set, generates ASCII in schema.

        :param unicode: Defaults to False: short-hand for the ``ucs2``
          character set, generates UNICODE in schema.

        :param binary: Defaults to False: short-hand, pick the binary
          collation type that matches the column's character set.  Generates
          BINARY in schema.  This does not affect the type of data stored,
          only the collation of character data.

        """
        kw.pop("strict", None)
        self._enum_init(enums, kw)
        _StringType.__init__(self, length=self.length, **kw)

    @classmethod
    def adapt_emulated_to_native(cls, impl, **kw):
        """Produce a MySQL native :class:`.mysql.ENUM` from plain
        :class:`.Enum`.

        """
        kw.setdefault("validate_strings", impl.validate_strings)
        kw.setdefault("values_callable", impl.values_callable)
        kw.setdefault("omit_aliases", impl._omit_aliases)
        return cls(**kw)

    def _object_value_for_elem(self, elem):
        # mysql sends back a blank string for any value that
        # was persisted that was not in the enums; that is, it does no
        # validation on the incoming data, it "truncates" it to be
        # the blank string.  Return it straight.
        if elem == "":
            return elem
        else:
            return super()._object_value_for_elem(elem)

    def __repr__(self):
        return util.generic_repr(
            self, to_inspect=[ENUM, _StringType, sqltypes.Enum]
        )


class SET(_StringType):
    """MySQL SET type."""

    __visit_name__ = "SET"

    def __init__(self, *values, **kw):
        """Construct a SET.

        E.g.::

          Column("myset", SET("foo", "bar", "baz"))

        The list of potential values is required in the case that this
        set will be used to generate DDL for a table, or if the
        :paramref:`.SET.retrieve_as_bitwise` flag is set to True.

        :param values: The range of valid values for this SET. The values
          are not quoted, they will be escaped and surrounded by single
          quotes when generating the schema.

        :param convert_unicode: Same flag as that of
         :paramref:`.String.convert_unicode`.

        :param collation: same as that of :paramref:`.String.collation`

        :param charset: same as that of :paramref:`.VARCHAR.charset`.

        :param ascii: same as that of :paramref:`.VARCHAR.ascii`.

        :param unicode: same as that of :paramref:`.VARCHAR.unicode`.

        :param binary: same as that of :paramref:`.VARCHAR.binary`.

        :param retrieve_as_bitwise: if True, the data for the set type will be
          persisted and selected using an integer value, where a set is coerced
          into a bitwise mask for persistence.  MySQL allows this mode which
          has the advantage of being able to store values unambiguously,
          such as the blank string ``''``.   The datatype will appear
          as the expression ``col + 0`` in a SELECT statement, so that the
          value is coerced into an integer value in result sets.
          This flag is required if one wishes
          to persist a set that can store the blank string ``''`` as a value.

          .. warning::

            When using :paramref:`.mysql.SET.retrieve_as_bitwise`, it is
            essential that the list of set values is expressed in the
            **exact same order** as exists on the MySQL database.

        """
        self.retrieve_as_bitwise = kw.pop("retrieve_as_bitwise", False)
        self.values = tuple(values)
        if not self.retrieve_as_bitwise and "" in values:
            raise exc.ArgumentError(
                "Can't use the blank value '' in a SET without "
                "setting retrieve_as_bitwise=True"
            )
        if self.retrieve_as_bitwise:
            self._bitmap = {
                value: 2**idx for idx, value in enumerate(self.values)
            }
            self._bitmap.update(
                (2**idx, value) for idx, value in enumerate(self.values)
            )
        length = max([len(v) for v in values] + [0])
        kw.setdefault("length", length)
        super().__init__(**kw)

    def column_expression(self, colexpr):
        if self.retrieve_as_bitwise:
            return sql.type_coerce(
                sql.type_coerce(colexpr, sqltypes.Integer) + 0, self
            )
        else:
            return colexpr

    def result_processor(self, dialect, coltype):
        if self.retrieve_as_bitwise:

            def process(value):
                if value is not None:
                    value = int(value)

                    return set(util.map_bits(self._bitmap.__getitem__, value))
                else:
                    return None

        else:
            super_convert = super().result_processor(dialect, coltype)

            def process(value):
                if isinstance(value, str):
                    # MySQLdb returns a string, let's parse
                    if super_convert:
                        value = super_convert(value)
                    return set(re.findall(r"[^,]+", value))
                else:
                    # mysql-connector-python does a naive
                    # split(",") which throws in an empty string
                    if value is not None:
                        value.discard("")
                    return value

        return process

    def bind_processor(self, dialect):
        super_convert = super().bind_processor(dialect)
        if self.retrieve_as_bitwise:

            def process(value):
                if value is None:
                    return None
                elif isinstance(value, (int, str)):
                    if super_convert:
                        return super_convert(value)
                    else:
                        return value
                else:
                    int_value = 0
                    for v in value:
                        int_value |= self._bitmap[v]
                    return int_value

        else:

            def process(value):
                # accept strings and int (actually bitflag) values directly
                if value is not None and not isinstance(value, (int, str)):
                    value = ",".join(value)

                if super_convert:
                    return super_convert(value)
                else:
                    return value

        return process

    def adapt(self, impltype, **kw):
        kw["retrieve_as_bitwise"] = self.retrieve_as_bitwise
        return util.constructor_copy(self, impltype, *self.values, **kw)

    def __repr__(self):
        return util.generic_repr(
            self,
            to_inspect=[SET, _StringType],
            additional_kw=[
                ("retrieve_as_bitwise", False),
            ],
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\quantum\boson.py ===
"""Bosonic quantum operators."""

from sympy.core.numbers import Integer
from sympy.core.singleton import S
from sympy.functions.elementary.complexes import conjugate
from sympy.functions.elementary.exponential import exp
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.physics.quantum import Operator
from sympy.physics.quantum import HilbertSpace, FockSpace, Ket, Bra
from sympy.functions.special.tensor_functions import KroneckerDelta


__all__ = [
    'BosonOp',
    'BosonFockKet',
    'BosonFockBra',
    'BosonCoherentKet',
    'BosonCoherentBra'
]


class BosonOp(Operator):
    """A bosonic operator that satisfies [a, Dagger(a)] == 1.

    Parameters
    ==========

    name : str
        A string that labels the bosonic mode.

    annihilation : bool
        A bool that indicates if the bosonic operator is an annihilation (True,
        default value) or creation operator (False)

    Examples
    ========

    >>> from sympy.physics.quantum import Dagger, Commutator
    >>> from sympy.physics.quantum.boson import BosonOp
    >>> a = BosonOp("a")
    >>> Commutator(a, Dagger(a)).doit()
    1
    """

    @property
    def name(self):
        return self.args[0]

    @property
    def is_annihilation(self):
        return bool(self.args[1])

    @classmethod
    def default_args(self):
        return ("a", True)

    def __new__(cls, *args, **hints):
        if not len(args) in [1, 2]:
            raise ValueError('1 or 2 parameters expected, got %s' % args)

        if len(args) == 1:
            args = (args[0], S.One)

        if len(args) == 2:
            args = (args[0], Integer(args[1]))

        return Operator.__new__(cls, *args)

    def _eval_commutator_BosonOp(self, other, **hints):
        if self.name == other.name:
            # [a^\dagger, a] = -1
            if not self.is_annihilation and other.is_annihilation:
                return S.NegativeOne

        elif 'independent' in hints and hints['independent']:
            # [a, b] = 0
            return S.Zero

        return None

    def _eval_commutator_FermionOp(self, other, **hints):
        return S.Zero

    def _eval_anticommutator_BosonOp(self, other, **hints):
        if 'independent' in hints and hints['independent']:
            # {a, b} = 2 * a * b, because [a, b] = 0
            return 2 * self * other

        return None

    def _eval_adjoint(self):
        return BosonOp(str(self.name), not self.is_annihilation)

    def _print_contents_latex(self, printer, *args):
        if self.is_annihilation:
            return r'{%s}' % str(self.name)
        else:
            return r'{{%s}^\dagger}' % str(self.name)

    def _print_contents(self, printer, *args):
        if self.is_annihilation:
            return r'%s' % str(self.name)
        else:
            return r'Dagger(%s)' % str(self.name)

    def _print_contents_pretty(self, printer, *args):
        from sympy.printing.pretty.stringpict import prettyForm
        pform = printer._print(self.args[0], *args)
        if self.is_annihilation:
            return pform
        else:
            return pform**prettyForm('\N{DAGGER}')


class BosonFockKet(Ket):
    """Fock state ket for a bosonic mode.

    Parameters
    ==========

    n : Number
        The Fock state number.

    """

    def __new__(cls, n):
        return Ket.__new__(cls, n)

    @property
    def n(self):
        return self.label[0]

    @classmethod
    def dual_class(self):
        return BosonFockBra

    @classmethod
    def _eval_hilbert_space(cls, label):
        return FockSpace()

    def _eval_innerproduct_BosonFockBra(self, bra, **hints):
        return KroneckerDelta(self.n, bra.n)

    def _apply_from_right_to_BosonOp(self, op, **options):
        if op.is_annihilation:
            return sqrt(self.n) * BosonFockKet(self.n - 1)
        else:
            return sqrt(self.n + 1) * BosonFockKet(self.n + 1)


class BosonFockBra(Bra):
    """Fock state bra for a bosonic mode.

    Parameters
    ==========

    n : Number
        The Fock state number.

    """

    def __new__(cls, n):
        return Bra.__new__(cls, n)

    @property
    def n(self):
        return self.label[0]

    @classmethod
    def dual_class(self):
        return BosonFockKet

    @classmethod
    def _eval_hilbert_space(cls, label):
        return FockSpace()


class BosonCoherentKet(Ket):
    """Coherent state ket for a bosonic mode.

    Parameters
    ==========

    alpha : Number, Symbol
        The complex amplitude of the coherent state.

    """

    def __new__(cls, alpha):
        return Ket.__new__(cls, alpha)

    @property
    def alpha(self):
        return self.label[0]

    @classmethod
    def dual_class(self):
        return BosonCoherentBra

    @classmethod
    def _eval_hilbert_space(cls, label):
        return HilbertSpace()

    def _eval_innerproduct_BosonCoherentBra(self, bra, **hints):
        if self.alpha == bra.alpha:
            return S.One
        else:
            return exp(-(abs(self.alpha)**2 + abs(bra.alpha)**2 - 2 * conjugate(bra.alpha) * self.alpha)/2)

    def _apply_from_right_to_BosonOp(self, op, **options):
        if op.is_annihilation:
            return self.alpha * self
        else:
            return None


class BosonCoherentBra(Bra):
    """Coherent state bra for a bosonic mode.

    Parameters
    ==========

    alpha : Number, Symbol
        The complex amplitude of the coherent state.

    """

    def __new__(cls, alpha):
        return Bra.__new__(cls, alpha)

    @property
    def alpha(self):
        return self.label[0]

    @classmethod
    def dual_class(self):
        return BosonCoherentKet

    def _apply_operator_BosonOp(self, op, **options):
        if not op.is_annihilation:
            return self.alpha * self
        else:
            return None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_core\_tests\test_thread_cache.py ===
from __future__ import annotations

import os
import threading
import time
from contextlib import contextmanager
from queue import Queue
from typing import TYPE_CHECKING, NoReturn

import pytest

from .. import _thread_cache
from .._thread_cache import ThreadCache, start_thread_soon
from .tutil import gc_collect_harder, slow

if TYPE_CHECKING:
    from collections.abc import Iterator

    from outcome import Outcome


def test_thread_cache_basics() -> None:
    q: Queue[Outcome[object]] = Queue()

    def fn() -> NoReturn:
        raise RuntimeError("hi")

    def deliver(outcome: Outcome[object]) -> None:
        q.put(outcome)

    start_thread_soon(fn, deliver)

    outcome = q.get()
    with pytest.raises(RuntimeError, match="hi"):
        outcome.unwrap()


def test_thread_cache_deref() -> None:
    res = [False]

    class del_me:
        def __call__(self) -> int:
            return 42

        def __del__(self) -> None:
            res[0] = True

    q: Queue[Outcome[int]] = Queue()

    def deliver(outcome: Outcome[int]) -> None:
        q.put(outcome)

    start_thread_soon(del_me(), deliver)
    outcome = q.get()
    assert outcome.unwrap() == 42

    gc_collect_harder()
    assert res[0]


@slow
def test_spawning_new_thread_from_deliver_reuses_starting_thread() -> None:
    # We know that no-one else is using the thread cache, so if we keep
    # submitting new jobs the instant the previous one is finished, we should
    # keep getting the same thread over and over. This tests both that the
    # thread cache is LIFO, and that threads can be assigned new work *before*
    # deliver exits.

    # Make sure there are a few threads running, so if we weren't LIFO then we
    # could grab the wrong one.
    q: Queue[Outcome[object]] = Queue()
    COUNT = 5
    for _ in range(COUNT):
        start_thread_soon(lambda: time.sleep(1), lambda result: q.put(result))
    for _ in range(COUNT):
        q.get().unwrap()

    seen_threads = set()
    done = threading.Event()

    def deliver(n: int, _: object) -> None:
        print(n)
        seen_threads.add(threading.current_thread())
        if n == 0:
            done.set()
        else:
            start_thread_soon(lambda: None, lambda _: deliver(n - 1, _))

    start_thread_soon(lambda: None, lambda _: deliver(5, _))

    done.wait()

    assert len(seen_threads) == 1


@slow
def test_idle_threads_exit(monkeypatch: pytest.MonkeyPatch) -> None:
    # Temporarily set the idle timeout to something tiny, to speed up the
    # test. (But non-zero, so that the worker loop will at least yield the
    # CPU.)
    monkeypatch.setattr(_thread_cache, "IDLE_TIMEOUT", 0.0001)

    q: Queue[threading.Thread] = Queue()
    start_thread_soon(lambda: None, lambda _: q.put(threading.current_thread()))
    seen_thread = q.get()
    # Since the idle timeout is 0, after sleeping for 1 second, the thread
    # should have exited
    time.sleep(1)
    assert not seen_thread.is_alive()


@contextmanager
def _join_started_threads() -> Iterator[None]:
    before = frozenset(threading.enumerate())
    try:
        yield
    finally:
        for thread in threading.enumerate():
            if thread not in before:
                thread.join(timeout=1.0)
                assert not thread.is_alive()


def test_race_between_idle_exit_and_job_assignment(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # This is a lock where the first few times you try to acquire it with a
    # timeout, it waits until the lock is available and then pretends to time
    # out. Using this in our thread cache implementation causes the following
    # sequence:
    #
    # 1. start_thread_soon grabs the worker thread, assigns it a job, and
    #    releases its lock.
    # 2. The worker thread wakes up (because the lock has been released), but
    #    the JankyLock lies to it and tells it that the lock timed out. So the
    #    worker thread tries to exit.
    # 3. The worker thread checks for the race between exiting and being
    #    assigned a job, and discovers that it *is* in the process of being
    #    assigned a job, so it loops around and tries to acquire the lock
    #    again.
    # 4. Eventually the JankyLock admits that the lock is available, and
    #    everything proceeds as normal.

    class JankyLock:
        def __init__(self) -> None:
            self._lock = threading.Lock()
            self._counter = 3

        def acquire(self, timeout: int = -1) -> bool:
            got_it = self._lock.acquire(timeout=timeout)
            if timeout == -1:
                return True
            elif got_it:
                if self._counter > 0:
                    self._counter -= 1
                    self._lock.release()
                    return False
                return True
            else:
                return False

        def release(self) -> None:
            self._lock.release()

    monkeypatch.setattr(_thread_cache, "Lock", JankyLock)

    with _join_started_threads():
        tc = ThreadCache()
        done = threading.Event()
        tc.start_thread_soon(lambda: None, lambda _: done.set())
        done.wait()
        # Let's kill the thread we started, so it doesn't hang around until the
        # test suite finishes. Doesn't really do any harm, but it can be confusing
        # to see it in debug output.
        monkeypatch.setattr(_thread_cache, "IDLE_TIMEOUT", 0.0001)
        tc.start_thread_soon(lambda: None, lambda _: None)


def test_raise_in_deliver(capfd: pytest.CaptureFixture[str]) -> None:
    seen_threads = set()

    def track_threads() -> None:
        seen_threads.add(threading.current_thread())

    def deliver(_: object) -> NoReturn:
        done.set()
        raise RuntimeError("don't do this")

    done = threading.Event()
    start_thread_soon(track_threads, deliver)
    done.wait()
    done = threading.Event()
    start_thread_soon(track_threads, lambda _: done.set())
    done.wait()
    assert len(seen_threads) == 1
    err = capfd.readouterr().err
    assert "don't do this" in err
    assert "delivering result" in err


@pytest.mark.skipif(not hasattr(os, "fork"), reason="os.fork isn't supported")
def test_clear_thread_cache_after_fork() -> None:
    assert hasattr(os, "fork")

    def foo() -> None:
        pass

    # ensure the thread cache exists
    done = threading.Event()
    start_thread_soon(foo, lambda _: done.set())
    done.wait()

    child_pid = os.fork()

    # try using it
    done = threading.Event()
    start_thread_soon(foo, lambda _: done.set())
    done.wait()

    if child_pid != 0:
        # if this test fails, this will hang, triggering a timeout.
        os.waitpid(child_pid, 0)
    else:
        # this is necessary because os._exit doesn't unwind the stack,
        # so coverage doesn't get to automatically stop and save
        # coverage information.
        try:
            import coverage

            cov = coverage.Coverage.current()
            # the following pragmas are necessary because if coverage:
            #  - isn't running, then it can't record the branch not
            #    taken
            #  - isn't installed, then it can't record the ImportError

            if cov:  # pragma: no branch
                cov.stop()
                cov.save()
        except ImportError:  # pragma: no cover
            pass

        os._exit(0)  # pragma: no cover  # coverage was stopped above.

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\ansi.py ===
import re
import sys
from contextlib import suppress
from typing import Iterable, NamedTuple, Optional

from .color import Color
from .style import Style
from .text import Text

re_ansi = re.compile(
    r"""
(?:\x1b[0-?])|
(?:\x1b\](.*?)\x1b\\)|
(?:\x1b([(@-Z\\-_]|\[[0-?]*[ -/]*[@-~]))
""",
    re.VERBOSE,
)


class _AnsiToken(NamedTuple):
    """Result of ansi tokenized string."""

    plain: str = ""
    sgr: Optional[str] = ""
    osc: Optional[str] = ""


def _ansi_tokenize(ansi_text: str) -> Iterable[_AnsiToken]:
    """Tokenize a string in to plain text and ANSI codes.

    Args:
        ansi_text (str): A String containing ANSI codes.

    Yields:
        AnsiToken: A named tuple of (plain, sgr, osc)
    """

    position = 0
    sgr: Optional[str]
    osc: Optional[str]
    for match in re_ansi.finditer(ansi_text):
        start, end = match.span(0)
        osc, sgr = match.groups()
        if start > position:
            yield _AnsiToken(ansi_text[position:start])
        if sgr:
            if sgr == "(":
                position = end + 1
                continue
            if sgr.endswith("m"):
                yield _AnsiToken("", sgr[1:-1], osc)
        else:
            yield _AnsiToken("", sgr, osc)
        position = end
    if position < len(ansi_text):
        yield _AnsiToken(ansi_text[position:])


SGR_STYLE_MAP = {
    1: "bold",
    2: "dim",
    3: "italic",
    4: "underline",
    5: "blink",
    6: "blink2",
    7: "reverse",
    8: "conceal",
    9: "strike",
    21: "underline2",
    22: "not dim not bold",
    23: "not italic",
    24: "not underline",
    25: "not blink",
    26: "not blink2",
    27: "not reverse",
    28: "not conceal",
    29: "not strike",
    30: "color(0)",
    31: "color(1)",
    32: "color(2)",
    33: "color(3)",
    34: "color(4)",
    35: "color(5)",
    36: "color(6)",
    37: "color(7)",
    39: "default",
    40: "on color(0)",
    41: "on color(1)",
    42: "on color(2)",
    43: "on color(3)",
    44: "on color(4)",
    45: "on color(5)",
    46: "on color(6)",
    47: "on color(7)",
    49: "on default",
    51: "frame",
    52: "encircle",
    53: "overline",
    54: "not frame not encircle",
    55: "not overline",
    90: "color(8)",
    91: "color(9)",
    92: "color(10)",
    93: "color(11)",
    94: "color(12)",
    95: "color(13)",
    96: "color(14)",
    97: "color(15)",
    100: "on color(8)",
    101: "on color(9)",
    102: "on color(10)",
    103: "on color(11)",
    104: "on color(12)",
    105: "on color(13)",
    106: "on color(14)",
    107: "on color(15)",
}


class AnsiDecoder:
    """Translate ANSI code in to styled Text."""

    def __init__(self) -> None:
        self.style = Style.null()

    def decode(self, terminal_text: str) -> Iterable[Text]:
        """Decode ANSI codes in an iterable of lines.

        Args:
            lines (Iterable[str]): An iterable of lines of terminal output.

        Yields:
            Text: Marked up Text.
        """
        for line in terminal_text.splitlines():
            yield self.decode_line(line)

    def decode_line(self, line: str) -> Text:
        """Decode a line containing ansi codes.

        Args:
            line (str): A line of terminal output.

        Returns:
            Text: A Text instance marked up according to ansi codes.
        """
        from_ansi = Color.from_ansi
        from_rgb = Color.from_rgb
        _Style = Style
        text = Text()
        append = text.append
        line = line.rsplit("\r", 1)[-1]
        for plain_text, sgr, osc in _ansi_tokenize(line):
            if plain_text:
                append(plain_text, self.style or None)
            elif osc is not None:
                if osc.startswith("8;"):
                    _params, semicolon, link = osc[2:].partition(";")
                    if semicolon:
                        self.style = self.style.update_link(link or None)
            elif sgr is not None:
                # Translate in to semi-colon separated codes
                # Ignore invalid codes, because we want to be lenient
                codes = [
                    min(255, int(_code) if _code else 0)
                    for _code in sgr.split(";")
                    if _code.isdigit() or _code == ""
                ]
                iter_codes = iter(codes)
                for code in iter_codes:
                    if code == 0:
                        # reset
                        self.style = _Style.null()
                    elif code in SGR_STYLE_MAP:
                        # styles
                        self.style += _Style.parse(SGR_STYLE_MAP[code])
                    elif code == 38:
                        #  Foreground
                        with suppress(StopIteration):
                            color_type = next(iter_codes)
                            if color_type == 5:
                                self.style += _Style.from_color(
                                    from_ansi(next(iter_codes))
                                )
                            elif color_type == 2:
                                self.style += _Style.from_color(
                                    from_rgb(
                                        next(iter_codes),
                                        next(iter_codes),
                                        next(iter_codes),
                                    )
                                )
                    elif code == 48:
                        # Background
                        with suppress(StopIteration):
                            color_type = next(iter_codes)
                            if color_type == 5:
                                self.style += _Style.from_color(
                                    None, from_ansi(next(iter_codes))
                                )
                            elif color_type == 2:
                                self.style += _Style.from_color(
                                    None,
                                    from_rgb(
                                        next(iter_codes),
                                        next(iter_codes),
                                        next(iter_codes),
                                    ),
                                )

        return text


if sys.platform != "win32" and __name__ == "__main__":  # pragma: no cover
    import io
    import os
    import pty
    import sys

    decoder = AnsiDecoder()

    stdout = io.BytesIO()

    def read(fd: int) -> bytes:
        data = os.read(fd, 1024)
        stdout.write(data)
        return data

    pty.spawn(sys.argv[1:], read)

    from .console import Console

    console = Console(record=True)

    stdout_result = stdout.getvalue().decode("utf-8")
    print(stdout_result)

    for line in decoder.decode(stdout_result):
        console.print(line)

    console.save_html("stdout.html")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_core\_asyncgens.py ===
from __future__ import annotations

import logging
import sys
import warnings
import weakref
from typing import TYPE_CHECKING, NoReturn, TypeVar

import attrs

from .. import _core
from .._util import name_asyncgen
from . import _run

# Used to log exceptions in async generator finalizers
ASYNCGEN_LOGGER = logging.getLogger("trio.async_generator_errors")

if TYPE_CHECKING:
    from collections.abc import Callable
    from types import AsyncGeneratorType

    from typing_extensions import ParamSpec

    _P = ParamSpec("_P")

    _WEAK_ASYNC_GEN_SET = weakref.WeakSet[AsyncGeneratorType[object, NoReturn]]
    _ASYNC_GEN_SET = set[AsyncGeneratorType[object, NoReturn]]
else:
    _WEAK_ASYNC_GEN_SET = weakref.WeakSet
    _ASYNC_GEN_SET = set

_R = TypeVar("_R")


@_core.disable_ki_protection
def _call_without_ki_protection(
    f: Callable[_P, _R],
    /,
    *args: _P.args,
    **kwargs: _P.kwargs,
) -> _R:
    return f(*args, **kwargs)


@attrs.define(eq=False)
class AsyncGenerators:
    # Async generators are added to this set when first iterated. Any
    # left after the main task exits will be closed before trio.run()
    # returns.  During most of the run, this is a WeakSet so GC works.
    # During shutdown, when we're finalizing all the remaining
    # asyncgens after the system nursery has been closed, it's a
    # regular set so we don't have to deal with GC firing at
    # unexpected times.
    alive: _WEAK_ASYNC_GEN_SET | _ASYNC_GEN_SET = attrs.Factory(_WEAK_ASYNC_GEN_SET)
    # The ids of foreign async generators are added to this set when first
    # iterated. Usually it is not safe to refer to ids like this, but because
    # we're using a finalizer we can ensure ids in this set do not outlive
    # their async generator.
    foreign: set[int] = attrs.Factory(set)

    # This collects async generators that get garbage collected during
    # the one-tick window between the system nursery closing and the
    # init task starting end-of-run asyncgen finalization.
    trailing_needs_finalize: _ASYNC_GEN_SET = attrs.Factory(_ASYNC_GEN_SET)

    prev_hooks: sys._asyncgen_hooks = attrs.field(init=False)

    def install_hooks(self, runner: _run.Runner) -> None:
        def firstiter(agen: AsyncGeneratorType[object, NoReturn]) -> None:
            if hasattr(_run.GLOBAL_RUN_CONTEXT, "task"):
                self.alive.add(agen)
            else:
                # An async generator first iterated outside of a Trio
                # task doesn't belong to Trio. Probably we're in guest
                # mode and the async generator belongs to our host.
                # A strong set of ids is one of the only good places to
                # remember this fact, at least until
                # https://github.com/python/cpython/issues/85093 is implemented.
                self.foreign.add(id(agen))
                if self.prev_hooks.firstiter is not None:
                    self.prev_hooks.firstiter(agen)

        def finalize_in_trio_context(
            agen: AsyncGeneratorType[object, NoReturn],
            agen_name: str,
        ) -> None:
            try:
                runner.spawn_system_task(
                    self._finalize_one,
                    agen,
                    agen_name,
                    name=f"close asyncgen {agen_name} (abandoned)",
                )
            except RuntimeError:
                # There is a one-tick window where the system nursery
                # is closed but the init task hasn't yet made
                # self.asyncgens a strong set to disable GC. We seem to
                # have hit it.
                self.trailing_needs_finalize.add(agen)

        @_core.enable_ki_protection
        def finalizer(agen: AsyncGeneratorType[object, NoReturn]) -> None:
            try:
                self.foreign.remove(id(agen))
            except KeyError:
                is_ours = True
            else:
                is_ours = False

            agen_name = name_asyncgen(agen)
            if is_ours:
                runner.entry_queue.run_sync_soon(
                    finalize_in_trio_context,
                    agen,
                    agen_name,
                )

                # Do this last, because it might raise an exception
                # depending on the user's warnings filter. (That
                # exception will be printed to the terminal and
                # ignored, since we're running in GC context.)
                warnings.warn(
                    f"Async generator {agen_name!r} was garbage collected before it "
                    "had been exhausted. Surround its use in 'async with "
                    "aclosing(...):' to ensure that it gets cleaned up as soon as "
                    "you're done using it.",
                    ResourceWarning,
                    stacklevel=2,
                    source=agen,
                )
            else:
                # Not ours -> forward to the host loop's async generator finalizer
                finalizer = self.prev_hooks.finalizer
                if finalizer is not None:
                    _call_without_ki_protection(finalizer, agen)
                else:
                    # Host has no finalizer.  Reimplement the default
                    # Python behavior with no hooks installed: throw in
                    # GeneratorExit, step once, raise RuntimeError if
                    # it doesn't exit.
                    closer = agen.aclose()
                    try:
                        # If the next thing is a yield, this will raise RuntimeError
                        # which we allow to propagate
                        _call_without_ki_protection(closer.send, None)
                    except StopIteration:
                        pass
                    else:
                        # If the next thing is an await, we get here. Give a nicer
                        # error than the default "async generator ignored GeneratorExit"
                        raise RuntimeError(
                            f"Non-Trio async generator {agen_name!r} awaited something "
                            "during finalization; install a finalization hook to "
                            "support this, or wrap it in 'async with aclosing(...):'",
                        )

        self.prev_hooks = sys.get_asyncgen_hooks()
        sys.set_asyncgen_hooks(firstiter=firstiter, finalizer=finalizer)  # type: ignore[arg-type]  # Finalizer doesn't use AsyncGeneratorType

    async def finalize_remaining(self, runner: _run.Runner) -> None:
        # This is called from init after shutting down the system nursery.
        # The only tasks running at this point are init and
        # the run_sync_soon task, and since the system nursery is closed,
        # there's no way for user code to spawn more.
        assert _core.current_task() is runner.init_task
        assert len(runner.tasks) == 2

        # To make async generator finalization easier to reason
        # about, we'll shut down asyncgen garbage collection by turning
        # the alive WeakSet into a regular set.
        self.alive = set(self.alive)

        # Process all pending run_sync_soon callbacks, in case one of
        # them was an asyncgen finalizer that snuck in under the wire.
        runner.entry_queue.run_sync_soon(runner.reschedule, runner.init_task)
        await _core.wait_task_rescheduled(
            lambda _: _core.Abort.FAILED,  # pragma: no cover
        )
        self.alive.update(self.trailing_needs_finalize)
        self.trailing_needs_finalize.clear()

        # None of the still-living tasks use async generators, so
        # every async generator must be suspended at a yield point --
        # there's no one to be doing the iteration. That's good,
        # because aclose() only works on an asyncgen that's suspended
        # at a yield point.  (If it's suspended at an event loop trap,
        # because someone is in the middle of iterating it, then you
        # get a RuntimeError on 3.8+, and a nasty surprise on earlier
        # versions due to https://bugs.python.org/issue32526.)
        #
        # However, once we start aclose() of one async generator, it
        # might start fetching the next value from another, thus
        # preventing us from closing that other (at least until
        # aclose() of the first one is complete).  This constraint
        # effectively requires us to finalize the remaining asyncgens
        # in arbitrary order, rather than doing all of them at the
        # same time. On 3.8+ we could defer any generator with
        # ag_running=True to a later batch, but that only catches
        # the case where our aclose() starts after the user's
        # asend()/etc. If our aclose() starts first, then the
        # user's asend()/etc will raise RuntimeError, since they're
        # probably not checking ag_running.
        #
        # It might be possible to allow some parallelized cleanup if
        # we can determine that a certain set of asyncgens have no
        # interdependencies, using gc.get_referents() and such.
        # But just doing one at a time will typically work well enough
        # (since each aclose() executes in a cancelled scope) and
        # is much easier to reason about.

        # It's possible that that cleanup code will itself create
        # more async generators, so we iterate repeatedly until
        # all are gone.
        while self.alive:
            batch = self.alive
            self.alive = _ASYNC_GEN_SET()
            for agen in batch:
                await self._finalize_one(agen, name_asyncgen(agen))

    def close(self) -> None:
        sys.set_asyncgen_hooks(*self.prev_hooks)

    async def _finalize_one(
        self,
        agen: AsyncGeneratorType[object, NoReturn],
        name: object,
    ) -> None:
        try:
            # This shield ensures that finalize_asyncgen never exits
            # with an exception, not even a Cancelled. The inside
            # is cancelled so there's no deadlock risk.
            with _core.CancelScope(shield=True) as cancel_scope:
                cancel_scope.cancel()
                await agen.aclose()
        except BaseException:
            ASYNCGEN_LOGGER.exception(
                "Exception ignored during finalization of async generator %r -- "
                "surround your use of the generator in 'async with aclosing(...):' "
                "to raise exceptions like this in the context where they're generated",
                name,
            )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\test_repl.py ===
from __future__ import annotations

import subprocess
import sys
from typing import Protocol

import pytest

import trio._repl


class RawInput(Protocol):
    def __call__(self, prompt: str = "") -> str: ...


def build_raw_input(cmds: list[str]) -> RawInput:
    """
    Pass in a list of strings.
    Returns a callable that returns each string, each time its called
    When there are not more strings to return, raise EOFError
    """
    cmds_iter = iter(cmds)
    prompts = []

    def _raw_helper(prompt: str = "") -> str:
        prompts.append(prompt)
        try:
            return next(cmds_iter)
        except StopIteration:
            raise EOFError from None

    return _raw_helper


def test_build_raw_input() -> None:
    """Quick test of our helper function."""
    raw_input = build_raw_input(["cmd1"])
    assert raw_input() == "cmd1"
    with pytest.raises(EOFError):
        raw_input()


# In 3.10 or later, types.FunctionType (used internally) will automatically
# attach __builtins__ to the function objects. However we need to explicitly
# include it for 3.9 support
def build_locals() -> dict[str, object]:
    return {"__builtins__": __builtins__}


async def test_basic_interaction(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """
    Run some basic commands through the interpreter while capturing stdout.
    Ensure that the interpreted prints the expected results.
    """
    console = trio._repl.TrioInteractiveConsole(repl_locals=build_locals())
    raw_input = build_raw_input(
        [
            # evaluate simple expression and recall the value
            "x = 1",
            "print(f'{x=}')",
            # Literal gets printed
            "'hello'",
            # define and call sync function
            "def func():",
            "  print(x + 1)",
            "",
            "func()",
            # define and call async function
            "async def afunc():",
            "  return 4",
            "",
            "await afunc()",
            # import works
            "import sys",
            "sys.stdout.write('hello stdout\\n')",
        ],
    )
    monkeypatch.setattr(console, "raw_input", raw_input)
    await trio._repl.run_repl(console)
    out, _err = capsys.readouterr()
    assert out.splitlines() == ["x=1", "'hello'", "2", "4", "hello stdout", "13"]


async def test_system_exits_quit_interpreter(monkeypatch: pytest.MonkeyPatch) -> None:
    console = trio._repl.TrioInteractiveConsole(repl_locals=build_locals())
    raw_input = build_raw_input(
        [
            "raise SystemExit",
        ],
    )
    monkeypatch.setattr(console, "raw_input", raw_input)
    with pytest.raises(SystemExit):
        await trio._repl.run_repl(console)


async def test_KI_interrupts(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    console = trio._repl.TrioInteractiveConsole(repl_locals=build_locals())
    raw_input = build_raw_input(
        [
            "import signal, trio, trio.lowlevel",
            "async def f():",
            "  trio.lowlevel.spawn_system_task("
            "    trio.to_thread.run_sync,"
            "    signal.raise_signal, signal.SIGINT,"
            "  )",  # just awaiting this kills the test runner?!
            "  await trio.sleep_forever()",
            "  print('should not see this')",
            "",
            "await f()",
            "print('AFTER KeyboardInterrupt')",
        ],
    )
    monkeypatch.setattr(console, "raw_input", raw_input)
    await trio._repl.run_repl(console)
    out, err = capsys.readouterr()
    assert "KeyboardInterrupt" in err
    assert "should" not in out
    assert "AFTER KeyboardInterrupt" in out


async def test_system_exits_in_exc_group(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    console = trio._repl.TrioInteractiveConsole(repl_locals=build_locals())
    raw_input = build_raw_input(
        [
            "import sys",
            "if sys.version_info < (3, 11):",
            "  from exceptiongroup import BaseExceptionGroup",
            "",
            "raise BaseExceptionGroup('', [RuntimeError(), SystemExit()])",
            "print('AFTER BaseExceptionGroup')",
        ],
    )
    monkeypatch.setattr(console, "raw_input", raw_input)
    await trio._repl.run_repl(console)
    out, _err = capsys.readouterr()
    # assert that raise SystemExit in an exception group
    # doesn't quit
    assert "AFTER BaseExceptionGroup" in out


async def test_system_exits_in_nested_exc_group(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    console = trio._repl.TrioInteractiveConsole(repl_locals=build_locals())
    raw_input = build_raw_input(
        [
            "import sys",
            "if sys.version_info < (3, 11):",
            "  from exceptiongroup import BaseExceptionGroup",
            "",
            "raise BaseExceptionGroup(",
            "  '', [BaseExceptionGroup('', [RuntimeError(), SystemExit()])])",
            "print('AFTER BaseExceptionGroup')",
        ],
    )
    monkeypatch.setattr(console, "raw_input", raw_input)
    await trio._repl.run_repl(console)
    out, _err = capsys.readouterr()
    # assert that raise SystemExit in an exception group
    # doesn't quit
    assert "AFTER BaseExceptionGroup" in out


async def test_base_exception_captured(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    console = trio._repl.TrioInteractiveConsole(repl_locals=build_locals())
    raw_input = build_raw_input(
        [
            # The statement after raise should still get executed
            "raise BaseException",
            "print('AFTER BaseException')",
        ],
    )
    monkeypatch.setattr(console, "raw_input", raw_input)
    await trio._repl.run_repl(console)
    out, err = capsys.readouterr()
    assert "_threads.py" not in err
    assert "_repl.py" not in err
    assert "AFTER BaseException" in out


async def test_exc_group_captured(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    console = trio._repl.TrioInteractiveConsole(repl_locals=build_locals())
    raw_input = build_raw_input(
        [
            # The statement after raise should still get executed
            "raise ExceptionGroup('', [KeyError()])",
            "print('AFTER ExceptionGroup')",
        ],
    )
    monkeypatch.setattr(console, "raw_input", raw_input)
    await trio._repl.run_repl(console)
    out, _err = capsys.readouterr()
    assert "AFTER ExceptionGroup" in out


async def test_base_exception_capture_from_coroutine(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    console = trio._repl.TrioInteractiveConsole(repl_locals=build_locals())
    raw_input = build_raw_input(
        [
            "async def async_func_raises_base_exception():",
            "  raise BaseException",
            "",
            # This will raise, but the statement after should still
            # be executed
            "await async_func_raises_base_exception()",
            "print('AFTER BaseException')",
        ],
    )
    monkeypatch.setattr(console, "raw_input", raw_input)
    await trio._repl.run_repl(console)
    out, err = capsys.readouterr()
    assert "_threads.py" not in err
    assert "_repl.py" not in err
    assert "AFTER BaseException" in out


def test_main_entrypoint() -> None:
    """
    Basic smoke test when running via the package __main__ entrypoint.
    """
    repl = subprocess.run([sys.executable, "-m", "trio"], input=b"exit()")
    assert repl.returncode == 0

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\mpmath\rational.py ===
import operator
import sys
from .libmp import int_types, mpf_hash, bitcount, from_man_exp, HASH_MODULUS

new = object.__new__

def create_reduced(p, q, _cache={}):
    key = p, q
    if key in _cache:
        return _cache[key]
    x, y = p, q
    while y:
        x, y = y, x % y
    if x != 1:
        p //= x
        q //= x
    v = new(mpq)
    v._mpq_ = p, q
    # Speedup integers, half-integers and other small fractions
    if q <= 4 and abs(key[0]) < 100:
        _cache[key] = v
    return v

class mpq(object):
    """
    Exact rational type, currently only intended for internal use.
    """

    __slots__ = ["_mpq_"]

    def __new__(cls, p, q=1):
        if type(p) is tuple:
            p, q = p
        elif hasattr(p, '_mpq_'):
            p, q = p._mpq_
        return create_reduced(p, q)

    def __repr__(s):
        return "mpq(%s,%s)" % s._mpq_

    def __str__(s):
        return "(%s/%s)" % s._mpq_

    def __int__(s):
        a, b = s._mpq_
        return a // b

    def __nonzero__(s):
        return bool(s._mpq_[0])

    __bool__ = __nonzero__

    def __hash__(s):
        a, b = s._mpq_
        if sys.version_info >= (3, 2):
            inverse = pow(b, HASH_MODULUS-2, HASH_MODULUS)
            if not inverse:
                h = sys.hash_info.inf
            else:
                h = (abs(a) * inverse) % HASH_MODULUS
            if a < 0: h = -h
            if h == -1: h = -2
            return h
        else:
            if b == 1:
                return hash(a)
            # Power of two: mpf compatible hash
            if not (b & (b-1)):
                return mpf_hash(from_man_exp(a, 1-bitcount(b)))
            return hash((a,b))

    def __eq__(s, t):
        ttype = type(t)
        if ttype is mpq:
            return s._mpq_ == t._mpq_
        if ttype in int_types:
            a, b = s._mpq_
            if b != 1:
                return False
            return a == t
        return NotImplemented

    def __ne__(s, t):
        ttype = type(t)
        if ttype is mpq:
            return s._mpq_ != t._mpq_
        if ttype in int_types:
            a, b = s._mpq_
            if b != 1:
                return True
            return a != t
        return NotImplemented

    def _cmp(s, t, op):
        ttype = type(t)
        if ttype in int_types:
            a, b = s._mpq_
            return op(a, t*b)
        if ttype is mpq:
            a, b = s._mpq_
            c, d = t._mpq_
            return op(a*d, b*c)
        return NotImplementedError

    def __lt__(s, t): return s._cmp(t, operator.lt)
    def __le__(s, t): return s._cmp(t, operator.le)
    def __gt__(s, t): return s._cmp(t, operator.gt)
    def __ge__(s, t): return s._cmp(t, operator.ge)

    def __abs__(s):
        a, b = s._mpq_
        if a >= 0:
            return s
        v = new(mpq)
        v._mpq_ = -a, b
        return v

    def __neg__(s):
        a, b = s._mpq_
        v = new(mpq)
        v._mpq_ = -a, b
        return v

    def __pos__(s):
        return s

    def __add__(s, t):
        ttype = type(t)
        if ttype is mpq:
            a, b = s._mpq_
            c, d = t._mpq_
            return create_reduced(a*d+b*c, b*d)
        if ttype in int_types:
            a, b = s._mpq_
            v = new(mpq)
            v._mpq_ = a+b*t, b
            return v
        return NotImplemented

    __radd__ = __add__

    def __sub__(s, t):
        ttype = type(t)
        if ttype is mpq:
            a, b = s._mpq_
            c, d = t._mpq_
            return create_reduced(a*d-b*c, b*d)
        if ttype in int_types:
            a, b = s._mpq_
            v = new(mpq)
            v._mpq_ = a-b*t, b
            return v
        return NotImplemented

    def __rsub__(s, t):
        ttype = type(t)
        if ttype is mpq:
            a, b = s._mpq_
            c, d = t._mpq_
            return create_reduced(b*c-a*d, b*d)
        if ttype in int_types:
            a, b = s._mpq_
            v = new(mpq)
            v._mpq_ = b*t-a, b
            return v
        return NotImplemented

    def __mul__(s, t):
        ttype = type(t)
        if ttype is mpq:
            a, b = s._mpq_
            c, d = t._mpq_
            return create_reduced(a*c, b*d)
        if ttype in int_types:
            a, b = s._mpq_
            return create_reduced(a*t, b)
        return NotImplemented

    __rmul__ = __mul__

    def __div__(s, t):
        ttype = type(t)
        if ttype is mpq:
            a, b = s._mpq_
            c, d = t._mpq_
            return create_reduced(a*d, b*c)
        if ttype in int_types:
            a, b = s._mpq_
            return create_reduced(a, b*t)
        return NotImplemented

    def __rdiv__(s, t):
        ttype = type(t)
        if ttype is mpq:
            a, b = s._mpq_
            c, d = t._mpq_
            return create_reduced(b*c, a*d)
        if ttype in int_types:
            a, b = s._mpq_
            return create_reduced(b*t, a)
        return NotImplemented

    def __pow__(s, t):
        ttype = type(t)
        if ttype in int_types:
            a, b = s._mpq_
            if t:
                if t < 0:
                    a, b, t = b, a, -t
                v = new(mpq)
                v._mpq_ = a**t, b**t
                return v
            raise ZeroDivisionError
        return NotImplemented


mpq_1 = mpq((1,1))
mpq_0 = mpq((0,1))
mpq_1_2 = mpq((1,2))
mpq_3_2 = mpq((3,2))
mpq_1_4 = mpq((1,4))
mpq_1_16 = mpq((1,16))
mpq_3_16 = mpq((3,16))
mpq_5_2 = mpq((5,2))
mpq_3_4 = mpq((3,4))
mpq_7_4 = mpq((7,4))
mpq_5_4 = mpq((5,4))


# Register with "numbers" ABC
#     We do not subclass, hence we do not use the @abstractmethod checks. While
#     this is less invasive it may turn out that we do not actually support
#     parts of the expected interfaces.  See
#     http://docs.python.org/2/library/numbers.html for list of abstract
#     methods.
try:
    import numbers
    numbers.Rational.register(mpq)
except ImportError:
    pass

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\descriptors\serialisable.py ===
# Copyright (c) 2010-2024 openpyxl

from copy import copy
from keyword import kwlist
KEYWORDS = frozenset(kwlist)

from . import Descriptor
from . import MetaSerialisable
from .sequence import (
    Sequence,
    NestedSequence,
    MultiSequencePart,
)
from .namespace import namespaced

from openpyxl.compat import safe_string
from openpyxl.xml.functions import (
    Element,
    localname,
)

seq_types = (list, tuple)

class Serialisable(metaclass=MetaSerialisable):
    """
    Objects can serialise to XML their attributes and child objects.
    The following class attributes are created by the metaclass at runtime:
    __attrs__ = attributes
    __nested__ = single-valued child treated as an attribute
    __elements__ = child elements
    """

    __attrs__ = None
    __nested__ = None
    __elements__ = None
    __namespaced__ = None

    idx_base = 0

    @property
    def tagname(self):
        raise(NotImplementedError)

    namespace = None

    @classmethod
    def from_tree(cls, node):
        """
        Create object from XML
        """
        # strip known namespaces from attributes
        attrib = dict(node.attrib)
        for key, ns in cls.__namespaced__:
            if ns in attrib:
                attrib[key] = attrib[ns]
                del attrib[ns]

        # strip attributes with unknown namespaces
        for key in list(attrib):
            if key.startswith('{'):
                del attrib[key]
            elif key in KEYWORDS:
                attrib["_" + key] = attrib[key]
                del attrib[key]
            elif "-" in key:
                n = key.replace("-", "_")
                attrib[n] = attrib[key]
                del attrib[key]

        if node.text and "attr_text" in cls.__attrs__:
            attrib["attr_text"] = node.text

        for el in node:
            tag = localname(el)
            if tag in KEYWORDS:
                tag = "_" + tag
            desc = getattr(cls, tag, None)
            if desc is None or isinstance(desc, property):
                continue

            if hasattr(desc, 'from_tree'):
                #descriptor manages conversion
                obj = desc.from_tree(el)
            else:
                if hasattr(desc.expected_type, "from_tree"):
                    #complex type
                    obj = desc.expected_type.from_tree(el)
                else:
                    #primitive
                    obj = el.text

            if isinstance(desc, NestedSequence):
                attrib[tag] = obj
            elif isinstance(desc, Sequence):
                attrib.setdefault(tag, [])
                attrib[tag].append(obj)
            elif isinstance(desc, MultiSequencePart):
                attrib.setdefault(desc.store, [])
                attrib[desc.store].append(obj)
            else:
                attrib[tag] = obj

        return cls(**attrib)


    def to_tree(self, tagname=None, idx=None, namespace=None):

        if tagname is None:
            tagname = self.tagname

        # keywords have to be masked
        if tagname.startswith("_"):
            tagname = tagname[1:]

        tagname = namespaced(self, tagname, namespace)
        namespace = getattr(self, "namespace", namespace)

        attrs = dict(self)
        for key, ns in self.__namespaced__:
            if key in attrs:
                attrs[ns] = attrs[key]
                del attrs[key]

        el = Element(tagname, attrs)
        if "attr_text" in self.__attrs__:
            el.text = safe_string(getattr(self, "attr_text"))

        for child_tag in self.__elements__:
            desc = getattr(self.__class__, child_tag, None)
            obj = getattr(self, child_tag)
            if hasattr(desc, "namespace") and hasattr(obj, 'namespace'):
                obj.namespace = desc.namespace

            if isinstance(obj, seq_types):
                if isinstance(desc, NestedSequence):
                    # wrap sequence in container
                    if not obj:
                        continue
                    nodes = [desc.to_tree(child_tag, obj, namespace)]
                elif isinstance(desc, Sequence):
                    # sequence
                    desc.idx_base = self.idx_base
                    nodes = (desc.to_tree(child_tag, obj, namespace))
                else: # property
                    nodes = (v.to_tree(child_tag, namespace) for v in obj)
                for node in nodes:
                    el.append(node)
            else:
                if child_tag in self.__nested__:
                    node = desc.to_tree(child_tag, obj, namespace)
                elif obj is None:
                    continue
                else:
                    node = obj.to_tree(child_tag)
                if node is not None:
                    el.append(node)
        return el


    def __iter__(self):
        for attr in self.__attrs__:
            value = getattr(self, attr)
            if attr.startswith("_"):
                attr = attr[1:]
            elif attr != "attr_text" and "_" in attr:
                desc = getattr(self.__class__, attr)
                if getattr(desc, "hyphenated", False):
                    attr = attr.replace("_", "-")
            if attr != "attr_text" and value is not None:
                yield attr, safe_string(value)


    def __eq__(self, other):
        if not self.__class__ == other.__class__:
            return False
        elif not dict(self) == dict(other):
            return False
        for el in self.__elements__:
            if getattr(self, el) != getattr(other, el):
                return False
        return True


    def __ne__(self, other):
        return not self == other


    def __repr__(self):
        s = u"<{0}.{1} object>\nParameters:".format(
            self.__module__,
            self.__class__.__name__
        )
        args = []
        for k in self.__attrs__ + self.__elements__:
            v = getattr(self, k)
            if isinstance(v, Descriptor):
                v = None
            args.append(u"{0}={1}".format(k, repr(v)))
        args = u", ".join(args)

        return u"\n".join([s, args])


    def __hash__(self):
        fields = []
        for attr in self.__attrs__ + self.__elements__:
            val = getattr(self, attr)
            if isinstance(val, list):
                val = tuple(val)
            fields.append(val)

        return hash(tuple(fields))


    def __add__(self, other):
        if type(self) != type(other):
            raise TypeError("Cannot combine instances of different types")
        vals = {}
        for attr in self.__attrs__:
            vals[attr] = getattr(self, attr) or getattr(other, attr)
        for el in self.__elements__:
            a = getattr(self, el)
            b = getattr(other, el)
            if a and b:
                vals[el] = a + b
            else:
                vals[el] = a or b
        return self.__class__(**vals)


    def __copy__(self):
        # serialise to xml and back to avoid shallow copies
        xml = self.to_tree(tagname="dummy")
        cp = self.__class__.from_tree(xml)
        # copy any non-persisted attributed
        for k in self.__dict__:
            if k not in self.__attrs__ + self.__elements__:
                v = copy(getattr(self, k))
                setattr(cp, k, v)
        return cp

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\utils\cell.py ===
# Copyright (c) 2010-2024 openpyxl

"""
Collection of utilities used within the package and also available for client code
"""
from functools import lru_cache
from itertools import chain, product
from string import ascii_uppercase, digits
import re

from .exceptions import CellCoordinatesException

# constants
COORD_RE = re.compile(r'^[$]?([A-Za-z]{1,3})[$]?(\d+)$')
COL_RANGE = """[A-Z]{1,3}:[A-Z]{1,3}:"""
ROW_RANGE = r"""\d+:\d+:"""
RANGE_EXPR = r"""
[$]?(?P<min_col>[A-Za-z]{1,3})?
[$]?(?P<min_row>\d+)?
(:[$]?(?P<max_col>[A-Za-z]{1,3})?
[$]?(?P<max_row>\d+)?)?
"""
ABSOLUTE_RE = re.compile('^' + RANGE_EXPR +'$', re.VERBOSE)
SHEET_TITLE = r"""
(('(?P<quoted>([^']|'')*)')|(?P<notquoted>[^'^ ^!]*))!"""
SHEETRANGE_RE = re.compile("""{0}(?P<cells>{1})(?=,?)""".format(
    SHEET_TITLE, RANGE_EXPR), re.VERBOSE)


def get_column_interval(start, end):
    """
    Given the start and end columns, return all the columns in the series.

    The start and end columns can be either column letters or 1-based
    indexes.
    """
    if isinstance(start, str):
        start = column_index_from_string(start)
    if isinstance(end, str):
        end = column_index_from_string(end)
    return [get_column_letter(x) for x in range(start, end + 1)]


def coordinate_from_string(coord_string):
    """Convert a coordinate string like 'B12' to a tuple ('B', 12)"""
    match = COORD_RE.match(coord_string)
    if not match:
        msg = f"Invalid cell coordinates ({coord_string})"
        raise CellCoordinatesException(msg)
    column, row = match.groups()
    row = int(row)
    if not row:
        msg = f"There is no row 0 ({coord_string})"
        raise CellCoordinatesException(msg)
    return column, row


def absolute_coordinate(coord_string):
    """Convert a coordinate to an absolute coordinate string (B12 -> $B$12)"""
    m = ABSOLUTE_RE.match(coord_string)
    if not m:
        raise ValueError(f"{coord_string} is not a valid coordinate range")

    d = m.groupdict('')
    for k, v in d.items():
        if v:
            d[k] = f"${v}"

    if d['max_col'] or d['max_row']:
        fmt = "{min_col}{min_row}:{max_col}{max_row}"
    else:
        fmt = "{min_col}{min_row}"
    return fmt.format(**d)


__decimal_to_alpha = [""] + list(ascii_uppercase)

@lru_cache(maxsize=None)
def get_column_letter(col_idx):
    """
    Convert decimal column position to its ASCII (base 26) form.

    Because column indices are 1-based, strides are actually pow(26, n) + 26
    Hence, a correction is applied between pow(26, n) and pow(26, 2) + 26 to
    prevent and additional column letter being prepended

    "A" == 1 == pow(26, 0)
    "Z" == 26 == pow(26, 0) + 26 // decimal equivalent 10
    "AA" == 27 == pow(26, 1) + 1
    "ZZ" == 702 == pow(26, 2) + 26 // decimal equivalent 100
    """

    if not 1 <= col_idx <= 18278:
        raise ValueError("Invalid column index {0}".format(col_idx))

    result = []

    if col_idx < 26:
        return __decimal_to_alpha[col_idx]

    while col_idx:
        col_idx, remainder = divmod(col_idx, 26)
        result.insert(0, __decimal_to_alpha[remainder])
        if not remainder:
            col_idx -= 1
            result.insert(0, "Z")

    return "".join(result)


__alpha_to_decimal = {letter:pos for pos, letter in enumerate(ascii_uppercase, 1)}
__powers = (1, 26, 676)

@lru_cache(maxsize=None)
def column_index_from_string(col):
    """
    Convert ASCII column name (base 26) to decimal with 1-based index

    Characters represent descending multiples of powers of 26

    "AFZ" == 26 * pow(26, 0) + 6 * pow(26, 1) + 1 * pow(26, 2)
    """
    error_msg = f"'{col}' is not a valid column name. Column names are from A to ZZZ"
    if len(col) > 3:
        raise ValueError(error_msg)
    idx = 0
    col = reversed(col.upper())
    for letter, power in zip(col, __powers):
        try:
            pos = __alpha_to_decimal[letter]
        except KeyError:
            raise ValueError(error_msg)
        idx += pos * power
    if not 0 < idx < 18279:
        raise ValueError(error_msg)
    return idx


def range_boundaries(range_string):
    """
    Convert a range string into a tuple of boundaries:
    (min_col, min_row, max_col, max_row)
    Cell coordinates will be converted into a range with the cell at both end
    """
    msg = "{0} is not a valid coordinate or range".format(range_string)
    m = ABSOLUTE_RE.match(range_string)
    if not m:
        raise ValueError(msg)

    min_col, min_row, sep, max_col, max_row = m.groups()

    if sep:
        cols = min_col, max_col
        rows = min_row, max_row

        if not (
            all(cols + rows) or
            all(cols) and not any(rows) or
            all(rows) and not any(cols)
        ):
            raise ValueError(msg)

    if min_col is not None:
        min_col = column_index_from_string(min_col)

    if min_row is not None:
        min_row = int(min_row)

    if max_col is not None:
        max_col = column_index_from_string(max_col)
    else:
        max_col = min_col

    if max_row is not None:
        max_row = int(max_row)
    else:
        max_row = min_row

    return min_col, min_row, max_col, max_row


def rows_from_range(range_string):
    """
    Get individual addresses for every cell in a range.
    Yields one row at a time.
    """
    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    rows = range(min_row, max_row + 1)
    cols = [get_column_letter(col) for col in range(min_col, max_col + 1)]
    for row in rows:
        yield tuple('{0}{1}'.format(col, row) for col in cols)


def cols_from_range(range_string):
    """
    Get individual addresses for every cell in a range.
    Yields one row at a time.
    """
    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    rows = range(min_row, max_row+1)
    cols = (get_column_letter(col) for col in range(min_col, max_col+1))
    for col in cols:
        yield tuple('{0}{1}'.format(col, row) for row in rows)


def coordinate_to_tuple(coordinate):
    """
    Convert an Excel style coordinate to (row, column) tuple
    """
    for idx, c in enumerate(coordinate):
        if c in digits:
            break
    col = coordinate[:idx]
    row = coordinate[idx:]
    return int(row), column_index_from_string(col)


def range_to_tuple(range_string):
    """
    Convert a worksheet range to the sheetname and maximum and minimum
    coordinate indices
    """
    m = SHEETRANGE_RE.match(range_string)
    if m is None:
        raise ValueError("Value must be of the form sheetname!A1:E4")
    sheetname = m.group("quoted") or m.group("notquoted")
    cells = m.group("cells")
    boundaries = range_boundaries(cells)
    return sheetname, boundaries


def quote_sheetname(sheetname):
    """
    Add quotes around sheetnames if they contain spaces.
    """
    if "'" in sheetname:
        sheetname = sheetname.replace("'", "''")

    sheetname = u"'{0}'".format(sheetname)
    return sheetname

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\quantization\matmul_bnb4_quantizer.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation. All rights reserved.
# Licensed under the MIT License. See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------

import argparse
import logging
import os

import numpy as np
import numpy.typing as npt
import onnx
from onnx.onnx_pb import GraphProto, ModelProto, NodeProto, TensorProto

from onnxruntime.capi._pybind_state import quantize_matmul_bnb4

from .onnx_model import ONNXModel
from .quant_utils import attribute_to_kwarg

logger = logging.getLogger(__name__)


class MatMulBnb4Quantizer:
    """Perform 4b quantization of constant MatMul weights using FP4 or NF4 data type"""

    ##################
    # quantization types, must be consistent with native code type
    # Bnb_DataType_t defined in blockwise_quant_block_bnb4.h

    # 4b floating point with bias of 3
    FP4 = 0

    # 4b NormalFloat
    NF4 = 1

    def __init__(self, model: ModelProto, quant_type: int, block_size: int, nodes_to_exclude=None):
        nodes_to_exclude = nodes_to_exclude or []
        assert quant_type in [MatMulBnb4Quantizer.FP4, MatMulBnb4Quantizer.NF4]
        self.model = ONNXModel(model)
        self.quant_type = quant_type
        self.block_size = block_size
        self.nodes_to_exclude = set(nodes_to_exclude)

    @staticmethod
    def __get_initializer(name, graph_path: list[GraphProto]) -> tuple[TensorProto, GraphProto]:
        for gid in range(len(graph_path) - 1, -1, -1):
            graph = graph_path[gid]
            for tensor in graph.initializer:
                if tensor.name == name:
                    return tensor, graph
        return None, None

    def bnb4_block_quant(self, fpweight: npt.ArrayLike) -> np.ndarray:
        """4b quantize fp32/fp16 weight"""

        if len(fpweight.shape) != 2:
            raise ValueError("Current bnb4 block quantization only supports 2D tensors!")
        # need to copy since the transposed weight still has the original memory layout
        # Linear4bit quantizes its weight data which is the transposed weight
        fpweight_t = fpweight.transpose().copy()

        rows, cols = fpweight.shape
        numel = rows * cols
        block_size = self.block_size
        num_blocks = (numel + block_size - 1) // block_size
        quantized_numel = (numel + 1) // 2

        packed = np.zeros(quantized_numel, dtype="uint8")
        absmax = np.zeros(num_blocks, dtype=fpweight.dtype)
        # block wise quantization, fpweight_t is flattened and divided into blocks
        quantize_matmul_bnb4(packed, fpweight_t, absmax, block_size, self.quant_type, cols, rows)

        return (packed, absmax)

    def _bnb4_matmul_node_weight(self, node: NodeProto, graph_stack: list[GraphProto]) -> NodeProto:
        """If the node is MatMul with fp32 const weight, quantize the weight with int4, and return the new node"""

        if node.op_type != "MatMul":
            return node  # only care about MatMul for now

        logger.debug(f"start to quantize {node.name} ...")
        if node.name in self.nodes_to_exclude:
            logger.debug(f"exclude to quantize {node.name} as specified by nodes_to_exclude...")
            return node

        inputB = node.input[1]  # noqa: N806
        B, Bs_graph = MatMulBnb4Quantizer.__get_initializer(inputB, graph_stack)  # noqa: N806
        if B is None:
            logger.debug("MatMul doesn't have const weight. Skip to quantize")
            return node  # only care about constant weight

        B_array = onnx.numpy_helper.to_array(B)  # noqa: N806
        if len(B_array.shape) != 2:
            logger.debug("MatMul weight is not 2D. Skip to quantize")
            return node  # can only process 2-D matrix

        packed, absmax = self.bnb4_block_quant(B_array)
        B_quant = onnx.numpy_helper.from_array(packed)  # noqa: N806
        B_quant.name = B.name + "_Bnb4"
        for input in Bs_graph.input:
            if input.name == inputB:
                Bs_graph.input.remove(input)
                break

        absmax_tensor = onnx.numpy_helper.from_array(absmax)
        absmax_tensor.name = B.name + "_absmax"

        Bs_graph.initializer.extend([B_quant, absmax_tensor])

        kwargs = {}
        rows, cols = B_array.shape
        kwargs["K"] = rows
        kwargs["N"] = cols
        kwargs["block_size"] = self.block_size
        kwargs["quant_type"] = self.quant_type

        matmul_bnb4_node = onnx.helper.make_node(
            "MatMulBnb4",
            inputs=[node.input[0], B_quant.name, absmax_tensor.name],
            outputs=[node.output[0]],
            name=node.name + "_Bnb4" if node.name else "",
            domain="com.microsoft",
            **kwargs,
        )

        logger.debug(f"complete quantization of {node.name} ...")

        return matmul_bnb4_node

    def _process_subgraph(self, graph_stack: list[GraphProto]):
        new_nodes = []
        graph = graph_stack[-1]

        for node in graph.node:
            graph_attrs = [
                attr
                for attr in node.attribute
                if attr.type == onnx.AttributeProto.GRAPH or attr.type == onnx.AttributeProto.GRAPHS
            ]
            if len(graph_attrs):
                kwargs = {}
                for attr in node.attribute:
                    if attr.type == onnx.AttributeProto.GRAPH:
                        # recursive call to take care of sub-graph
                        graph_stack.append(attr.g)
                        kv = {attr.name: self._process_subgraph(graph_stack)}
                    elif attr.type == onnx.AttributeProto.GRAPHS:
                        value = []
                        for subgraph in attr.graphs:
                            # recursive call to take care of sub-graph
                            graph_stack.append(subgraph)
                            value.extend([self._process_subgraph(graph_stack)])
                        kv = {attr.name: value}
                    else:
                        kv = attribute_to_kwarg(attr)
                    kwargs.update(kv)
                node = onnx.helper.make_node(  # noqa: PLW2901
                    node.op_type, node.input, node.output, name=node.name, **kwargs
                )

            new_nodes.append(self._bnb4_matmul_node_weight(node, graph_stack))

        graph.ClearField("node")
        graph.node.extend(new_nodes)
        graph_stack.pop()
        return graph

    def process(self):
        # use a stack to keep track of sub-graphs
        graph_stack = [self.model.graph()]
        opset_import = self.model.opset_import()

        has_ms_domain = False
        for opset in opset_import:
            if opset.domain == "com.microsoft":
                has_ms_domain = True
        if not has_ms_domain:
            opset_import.extend([onnx.helper.make_opsetid("com.microsoft", 1)])

        self._process_subgraph(graph_stack)
        self.model.clean_initializers()


def parse_args():
    parser = argparse.ArgumentParser(
        description="""Blockwise FP4/NF4 quantization for MatMul 2D weight matrices.

A weight matrix is partitioned into blocks, where each block is a contiguous
subset inside the flattened transposed weight matrix. Each block is quantized
into a set of 4b integers with an absolute value scaling factor.
"""
    )

    parser.add_argument("--input_model", required=True, help="Path to the input model file")
    parser.add_argument("--output_model", required=True, help="Path to the output model file")
    parser.add_argument(
        "--quant_type",
        required=False,
        default=1,
        choices=[MatMulBnb4Quantizer.FP4, MatMulBnb4Quantizer.NF4],
        help="Quantization data type. 0: FP4, 1: NF4",
    )
    parser.add_argument(
        "--block_size",
        required=False,
        default=64,
        help="Block size for blockwise quantization. Note: bnb.nn.Linear4bit only uses block_size=64",
    )
    parser.add_argument("-v", "--verbose", required=False, action="store_true")
    parser.set_defaults(verbose=False)
    parser.add_argument(
        "--nodes_to_exclude",
        nargs="+",
        type=str,
        required=False,
        default=[],
        help="Specify the nodes to be excluded from quantization with node names",
    )

    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    if args.verbose:
        logger.setLevel(logging.DEBUG)

    input_model_path = args.input_model
    output_model_path = args.output_model

    if os.path.exists(output_model_path):
        logger.error(f"file {output_model_path} already exists")
        raise Exception(f"file {output_model_path} already exists")

    model = onnx.load(input_model_path)
    quant = MatMulBnb4Quantizer(model, args.quant_type, args.block_size, nodes_to_exclude=args.nodes_to_exclude)
    quant.process()
    quant.model.save_model_to_file(output_model_path, True)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\outcome\_impl.py ===
from __future__ import annotations

import abc
from typing import (
    TYPE_CHECKING,
    AsyncGenerator,
    Awaitable,
    Callable,
    Generator,
    Generic,
    NoReturn,
    TypeVar,
    Union,
    overload,
)

import attr

from ._util import AlreadyUsedError, remove_tb_frames

if TYPE_CHECKING:
    from typing_extensions import ParamSpec, final
    ArgsT = ParamSpec("ArgsT")
else:

    def final(func):
        return func


__all__ = ['Error', 'Outcome', 'Maybe', 'Value', 'acapture', 'capture']

ValueT = TypeVar("ValueT", covariant=True)
ResultT = TypeVar("ResultT")


@overload
def capture(
        # NoReturn = raises exception, so we should get an error.
        sync_fn: Callable[ArgsT, NoReturn],
        *args: ArgsT.args,
        **kwargs: ArgsT.kwargs,
) -> Error:
    ...


@overload
def capture(
        sync_fn: Callable[ArgsT, ResultT],
        *args: ArgsT.args,
        **kwargs: ArgsT.kwargs,
) -> Value[ResultT] | Error:
    ...


def capture(
        sync_fn: Callable[ArgsT, ResultT],
        *args: ArgsT.args,
        **kwargs: ArgsT.kwargs,
) -> Value[ResultT] | Error:
    """Run ``sync_fn(*args, **kwargs)`` and capture the result.

    Returns:
      Either a :class:`Value` or :class:`Error` as appropriate.

    """
    try:
        return Value(sync_fn(*args, **kwargs))
    except BaseException as exc:
        exc = remove_tb_frames(exc, 1)
        return Error(exc)


@overload
async def acapture(
        async_fn: Callable[ArgsT, Awaitable[NoReturn]],
        *args: ArgsT.args,
        **kwargs: ArgsT.kwargs,
) -> Error:
    ...


@overload
async def acapture(
        async_fn: Callable[ArgsT, Awaitable[ResultT]],
        *args: ArgsT.args,
        **kwargs: ArgsT.kwargs,
) -> Value[ResultT] | Error:
    ...


async def acapture(
        async_fn: Callable[ArgsT, Awaitable[ResultT]],
        *args: ArgsT.args,
        **kwargs: ArgsT.kwargs,
) -> Value[ResultT] | Error:
    """Run ``await async_fn(*args, **kwargs)`` and capture the result.

    Returns:
      Either a :class:`Value` or :class:`Error` as appropriate.

    """
    try:
        return Value(await async_fn(*args, **kwargs))
    except BaseException as exc:
        exc = remove_tb_frames(exc, 1)
        return Error(exc)


@attr.s(repr=False, init=False, slots=True)
class Outcome(abc.ABC, Generic[ValueT]):
    """An abstract class representing the result of a Python computation.

    This class has two concrete subclasses: :class:`Value` representing a
    value, and :class:`Error` representing an exception.

    In addition to the methods described below, comparison operators on
    :class:`Value` and :class:`Error` objects (``==``, ``<``, etc.) check that
    the other object is also a :class:`Value` or :class:`Error` object
    respectively, and then compare the contained objects.

    :class:`Outcome` objects are hashable if the contained objects are
    hashable.

    """
    _unwrapped: bool = attr.ib(default=False, eq=False, init=False)

    def _set_unwrapped(self) -> None:
        if self._unwrapped:
            raise AlreadyUsedError
        object.__setattr__(self, '_unwrapped', True)

    @abc.abstractmethod
    def unwrap(self) -> ValueT:
        """Return or raise the contained value or exception.

        These two lines of code are equivalent::

           x = fn(*args)
           x = outcome.capture(fn, *args).unwrap()

        """

    @abc.abstractmethod
    def send(self, gen: Generator[ResultT, ValueT, object]) -> ResultT:
        """Send or throw the contained value or exception into the given
        generator object.

        Args:
          gen: A generator object supporting ``.send()`` and ``.throw()``
              methods.

        """

    @abc.abstractmethod
    async def asend(self, agen: AsyncGenerator[ResultT, ValueT]) -> ResultT:
        """Send or throw the contained value or exception into the given async
        generator object.

        Args:
          agen: An async generator object supporting ``.asend()`` and
              ``.athrow()`` methods.

        """


@final
@attr.s(frozen=True, repr=False, slots=True)
class Value(Outcome[ValueT], Generic[ValueT]):
    """Concrete :class:`Outcome` subclass representing a regular value.

    """

    value: ValueT = attr.ib()
    """The contained value."""

    def __repr__(self) -> str:
        return f'Value({self.value!r})'

    def unwrap(self) -> ValueT:
        self._set_unwrapped()
        return self.value

    def send(self, gen: Generator[ResultT, ValueT, object]) -> ResultT:
        self._set_unwrapped()
        return gen.send(self.value)

    async def asend(self, agen: AsyncGenerator[ResultT, ValueT]) -> ResultT:
        self._set_unwrapped()
        return await agen.asend(self.value)


@final
@attr.s(frozen=True, repr=False, slots=True)
class Error(Outcome[NoReturn]):
    """Concrete :class:`Outcome` subclass representing a raised exception.

    """

    error: BaseException = attr.ib(
        validator=attr.validators.instance_of(BaseException)
    )
    """The contained exception object."""

    def __repr__(self) -> str:
        return f'Error({self.error!r})'

    def unwrap(self) -> NoReturn:
        self._set_unwrapped()
        # Tracebacks show the 'raise' line below out of context, so let's give
        # this variable a name that makes sense out of context.
        captured_error = self.error
        try:
            raise captured_error
        finally:
            # We want to avoid creating a reference cycle here. Python does
            # collect cycles just fine, so it wouldn't be the end of the world
            # if we did create a cycle, but the cyclic garbage collector adds
            # latency to Python programs, and the more cycles you create, the
            # more often it runs, so it's nicer to avoid creating them in the
            # first place. For more details see:
            #
            #    https://github.com/python-trio/trio/issues/1770
            #
            # In particuar, by deleting this local variables from the 'unwrap'
            # methods frame, we avoid the 'captured_error' object's
            # __traceback__ from indirectly referencing 'captured_error'.
            del captured_error, self

    def send(self, gen: Generator[ResultT, NoReturn, object]) -> ResultT:
        self._set_unwrapped()
        return gen.throw(self.error)

    async def asend(self, agen: AsyncGenerator[ResultT, NoReturn]) -> ResultT:
        self._set_unwrapped()
        return await agen.athrow(self.error)


# A convenience alias to a union of both results, allowing exhaustiveness checking.
Maybe = Union[Value[ValueT], Error]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\_numba\executor.py ===
from __future__ import annotations

import functools
from typing import (
    TYPE_CHECKING,
    Any,
    Callable,
)

if TYPE_CHECKING:
    from pandas._typing import Scalar

import numpy as np

from pandas.compat._optional import import_optional_dependency


@functools.cache
def generate_apply_looper(func, nopython=True, nogil=True, parallel=False):
    if TYPE_CHECKING:
        import numba
    else:
        numba = import_optional_dependency("numba")
    nb_compat_func = numba.extending.register_jitable(func)

    @numba.jit(nopython=nopython, nogil=nogil, parallel=parallel)
    def nb_looper(values, axis):
        # Operate on the first row/col in order to get
        # the output shape
        if axis == 0:
            first_elem = values[:, 0]
            dim0 = values.shape[1]
        else:
            first_elem = values[0]
            dim0 = values.shape[0]
        res0 = nb_compat_func(first_elem)
        # Use np.asarray to get shape for
        # https://github.com/numba/numba/issues/4202#issuecomment-1185981507
        buf_shape = (dim0,) + np.atleast_1d(np.asarray(res0)).shape
        if axis == 0:
            buf_shape = buf_shape[::-1]
        buff = np.empty(buf_shape)

        if axis == 1:
            buff[0] = res0
            for i in numba.prange(1, values.shape[0]):
                buff[i] = nb_compat_func(values[i])
        else:
            buff[:, 0] = res0
            for j in numba.prange(1, values.shape[1]):
                buff[:, j] = nb_compat_func(values[:, j])
        return buff

    return nb_looper


@functools.cache
def make_looper(func, result_dtype, is_grouped_kernel, nopython, nogil, parallel):
    if TYPE_CHECKING:
        import numba
    else:
        numba = import_optional_dependency("numba")

    if is_grouped_kernel:

        @numba.jit(nopython=nopython, nogil=nogil, parallel=parallel)
        def column_looper(
            values: np.ndarray,
            labels: np.ndarray,
            ngroups: int,
            min_periods: int,
            *args,
        ):
            result = np.empty((values.shape[0], ngroups), dtype=result_dtype)
            na_positions = {}
            for i in numba.prange(values.shape[0]):
                output, na_pos = func(
                    values[i], result_dtype, labels, ngroups, min_periods, *args
                )
                result[i] = output
                if len(na_pos) > 0:
                    na_positions[i] = np.array(na_pos)
            return result, na_positions

    else:

        @numba.jit(nopython=nopython, nogil=nogil, parallel=parallel)
        def column_looper(
            values: np.ndarray,
            start: np.ndarray,
            end: np.ndarray,
            min_periods: int,
            *args,
        ):
            result = np.empty((values.shape[0], len(start)), dtype=result_dtype)
            na_positions = {}
            for i in numba.prange(values.shape[0]):
                output, na_pos = func(
                    values[i], result_dtype, start, end, min_periods, *args
                )
                result[i] = output
                if len(na_pos) > 0:
                    na_positions[i] = np.array(na_pos)
            return result, na_positions

    return column_looper


default_dtype_mapping: dict[np.dtype, Any] = {
    np.dtype("int8"): np.int64,
    np.dtype("int16"): np.int64,
    np.dtype("int32"): np.int64,
    np.dtype("int64"): np.int64,
    np.dtype("uint8"): np.uint64,
    np.dtype("uint16"): np.uint64,
    np.dtype("uint32"): np.uint64,
    np.dtype("uint64"): np.uint64,
    np.dtype("float32"): np.float64,
    np.dtype("float64"): np.float64,
    np.dtype("complex64"): np.complex128,
    np.dtype("complex128"): np.complex128,
}


# TODO: Preserve complex dtypes

float_dtype_mapping: dict[np.dtype, Any] = {
    np.dtype("int8"): np.float64,
    np.dtype("int16"): np.float64,
    np.dtype("int32"): np.float64,
    np.dtype("int64"): np.float64,
    np.dtype("uint8"): np.float64,
    np.dtype("uint16"): np.float64,
    np.dtype("uint32"): np.float64,
    np.dtype("uint64"): np.float64,
    np.dtype("float32"): np.float64,
    np.dtype("float64"): np.float64,
    np.dtype("complex64"): np.float64,
    np.dtype("complex128"): np.float64,
}

identity_dtype_mapping: dict[np.dtype, Any] = {
    np.dtype("int8"): np.int8,
    np.dtype("int16"): np.int16,
    np.dtype("int32"): np.int32,
    np.dtype("int64"): np.int64,
    np.dtype("uint8"): np.uint8,
    np.dtype("uint16"): np.uint16,
    np.dtype("uint32"): np.uint32,
    np.dtype("uint64"): np.uint64,
    np.dtype("float32"): np.float32,
    np.dtype("float64"): np.float64,
    np.dtype("complex64"): np.complex64,
    np.dtype("complex128"): np.complex128,
}


def generate_shared_aggregator(
    func: Callable[..., Scalar],
    dtype_mapping: dict[np.dtype, np.dtype],
    is_grouped_kernel: bool,
    nopython: bool,
    nogil: bool,
    parallel: bool,
):
    """
    Generate a Numba function that loops over the columns 2D object and applies
    a 1D numba kernel over each column.

    Parameters
    ----------
    func : function
        aggregation function to be applied to each column
    dtype_mapping: dict or None
        If not None, maps a dtype to a result dtype.
        Otherwise, will fall back to default mapping.
    is_grouped_kernel: bool, default False
        Whether func operates using the group labels (True)
        or using starts/ends arrays

        If true, you also need to pass the number of groups to this function
    nopython : bool
        nopython to be passed into numba.jit
    nogil : bool
        nogil to be passed into numba.jit
    parallel : bool
        parallel to be passed into numba.jit

    Returns
    -------
    Numba function
    """

    # A wrapper around the looper function,
    # to dispatch based on dtype since numba is unable to do that in nopython mode

    # It also post-processes the values by inserting nans where number of observations
    # is less than min_periods
    # Cannot do this in numba nopython mode
    # (you'll run into type-unification error when you cast int -> float)
    def looper_wrapper(
        values,
        start=None,
        end=None,
        labels=None,
        ngroups=None,
        min_periods: int = 0,
        **kwargs,
    ):
        result_dtype = dtype_mapping[values.dtype]
        column_looper = make_looper(
            func, result_dtype, is_grouped_kernel, nopython, nogil, parallel
        )
        # Need to unpack kwargs since numba only supports *args
        if is_grouped_kernel:
            result, na_positions = column_looper(
                values, labels, ngroups, min_periods, *kwargs.values()
            )
        else:
            result, na_positions = column_looper(
                values, start, end, min_periods, *kwargs.values()
            )
        if result.dtype.kind == "i":
            # Look if na_positions is not empty
            # If so, convert the whole block
            # This is OK since int dtype cannot hold nan,
            # so if min_periods not satisfied for 1 col, it is not satisfied for
            # all columns at that index
            for na_pos in na_positions.values():
                if len(na_pos) > 0:
                    result = result.astype("float64")
                    break
        # TODO: Optimize this
        for i, na_pos in na_positions.items():
            if len(na_pos) > 0:
                result[i, na_pos] = np.nan
        return result

    return looper_wrapper

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\orm\exc.py ===
# orm/exc.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

"""SQLAlchemy ORM exceptions."""

from __future__ import annotations

from typing import Any
from typing import Optional
from typing import Tuple
from typing import Type
from typing import TYPE_CHECKING
from typing import TypeVar

from .util import _mapper_property_as_plain_name
from .. import exc as sa_exc
from .. import util
from ..exc import MultipleResultsFound  # noqa
from ..exc import NoResultFound  # noqa

if TYPE_CHECKING:
    from .interfaces import LoaderStrategy
    from .interfaces import MapperProperty
    from .state import InstanceState

_T = TypeVar("_T", bound=Any)

NO_STATE = (AttributeError, KeyError)
"""Exception types that may be raised by instrumentation implementations."""


class StaleDataError(sa_exc.SQLAlchemyError):
    """An operation encountered database state that is unaccounted for.

    Conditions which cause this to happen include:

    * A flush may have attempted to update or delete rows
      and an unexpected number of rows were matched during
      the UPDATE or DELETE statement.   Note that when
      version_id_col is used, rows in UPDATE or DELETE statements
      are also matched against the current known version
      identifier.

    * A mapped object with version_id_col was refreshed,
      and the version number coming back from the database does
      not match that of the object itself.

    * A object is detached from its parent object, however
      the object was previously attached to a different parent
      identity which was garbage collected, and a decision
      cannot be made if the new parent was really the most
      recent "parent".

    """


ConcurrentModificationError = StaleDataError


class FlushError(sa_exc.SQLAlchemyError):
    """A invalid condition was detected during flush()."""


class MappedAnnotationError(sa_exc.ArgumentError):
    """Raised when ORM annotated declarative cannot interpret the
    expression present inside of the :class:`.Mapped` construct.

    .. versionadded:: 2.0.40

    """


class UnmappedError(sa_exc.InvalidRequestError):
    """Base for exceptions that involve expected mappings not present."""


class ObjectDereferencedError(sa_exc.SQLAlchemyError):
    """An operation cannot complete due to an object being garbage
    collected.

    """


class DetachedInstanceError(sa_exc.SQLAlchemyError):
    """An attempt to access unloaded attributes on a
    mapped instance that is detached."""

    code = "bhk3"


class UnmappedInstanceError(UnmappedError):
    """An mapping operation was requested for an unknown instance."""

    @util.preload_module("sqlalchemy.orm.base")
    def __init__(self, obj: object, msg: Optional[str] = None):
        base = util.preloaded.orm_base

        if not msg:
            try:
                base.class_mapper(type(obj))
                name = _safe_cls_name(type(obj))
                msg = (
                    "Class %r is mapped, but this instance lacks "
                    "instrumentation.  This occurs when the instance "
                    "is created before sqlalchemy.orm.mapper(%s) "
                    "was called." % (name, name)
                )
            except UnmappedClassError:
                msg = f"Class '{_safe_cls_name(type(obj))}' is not mapped"
                if isinstance(obj, type):
                    msg += (
                        "; was a class (%s) supplied where an instance was "
                        "required?" % _safe_cls_name(obj)
                    )
        UnmappedError.__init__(self, msg)

    def __reduce__(self) -> Any:
        return self.__class__, (None, self.args[0])


class UnmappedClassError(UnmappedError):
    """An mapping operation was requested for an unknown class."""

    def __init__(self, cls: Type[_T], msg: Optional[str] = None):
        if not msg:
            msg = _default_unmapped(cls)
        UnmappedError.__init__(self, msg)

    def __reduce__(self) -> Any:
        return self.__class__, (None, self.args[0])


class ObjectDeletedError(sa_exc.InvalidRequestError):
    """A refresh operation failed to retrieve the database
    row corresponding to an object's known primary key identity.

    A refresh operation proceeds when an expired attribute is
    accessed on an object, or when :meth:`_query.Query.get` is
    used to retrieve an object which is, upon retrieval, detected
    as expired.   A SELECT is emitted for the target row
    based on primary key; if no row is returned, this
    exception is raised.

    The true meaning of this exception is simply that
    no row exists for the primary key identifier associated
    with a persistent object.   The row may have been
    deleted, or in some cases the primary key updated
    to a new value, outside of the ORM's management of the target
    object.

    """

    @util.preload_module("sqlalchemy.orm.base")
    def __init__(self, state: InstanceState[Any], msg: Optional[str] = None):
        base = util.preloaded.orm_base

        if not msg:
            msg = (
                "Instance '%s' has been deleted, or its "
                "row is otherwise not present." % base.state_str(state)
            )

        sa_exc.InvalidRequestError.__init__(self, msg)

    def __reduce__(self) -> Any:
        return self.__class__, (None, self.args[0])


class UnmappedColumnError(sa_exc.InvalidRequestError):
    """Mapping operation was requested on an unknown column."""


class LoaderStrategyException(sa_exc.InvalidRequestError):
    """A loader strategy for an attribute does not exist."""

    def __init__(
        self,
        applied_to_property_type: Type[Any],
        requesting_property: MapperProperty[Any],
        applies_to: Optional[Type[MapperProperty[Any]]],
        actual_strategy_type: Optional[Type[LoaderStrategy]],
        strategy_key: Tuple[Any, ...],
    ):
        if actual_strategy_type is None:
            sa_exc.InvalidRequestError.__init__(
                self,
                "Can't find strategy %s for %s"
                % (strategy_key, requesting_property),
            )
        else:
            assert applies_to is not None
            sa_exc.InvalidRequestError.__init__(
                self,
                'Can\'t apply "%s" strategy to property "%s", '
                'which is a "%s"; this loader strategy is intended '
                'to be used with a "%s".'
                % (
                    util.clsname_as_plain_name(actual_strategy_type),
                    requesting_property,
                    _mapper_property_as_plain_name(applied_to_property_type),
                    _mapper_property_as_plain_name(applies_to),
                ),
            )


def _safe_cls_name(cls: Type[Any]) -> str:
    cls_name: Optional[str]
    try:
        cls_name = ".".join((cls.__module__, cls.__name__))
    except AttributeError:
        cls_name = getattr(cls, "__name__", None)
        if cls_name is None:
            cls_name = repr(cls)
    return cls_name


@util.preload_module("sqlalchemy.orm.base")
def _default_unmapped(cls: Type[Any]) -> Optional[str]:
    base = util.preloaded.orm_base

    try:
        mappers = base.manager_of_class(cls).mappers  # type: ignore
    except (
        UnmappedClassError,
        TypeError,
    ) + NO_STATE:
        mappers = {}
    name = _safe_cls_name(cls)

    if not mappers:
        return f"Class '{name}' is not mapped"
    else:
        return None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\domains\modularinteger.py ===
"""Implementation of :class:`ModularInteger` class. """

from __future__ import annotations
from typing import Any

import operator

from sympy.polys.polyutils import PicklableWithSlots
from sympy.polys.polyerrors import CoercionFailed
from sympy.polys.domains.domainelement import DomainElement

from sympy.utilities import public
from sympy.utilities.exceptions import sympy_deprecation_warning

@public
class ModularInteger(PicklableWithSlots, DomainElement):
    """A class representing a modular integer. """

    mod, dom, sym, _parent = None, None, None, None

    __slots__ = ('val',)

    def parent(self):
        return self._parent

    def __init__(self, val):
        if isinstance(val, self.__class__):
            self.val = val.val % self.mod
        else:
            self.val = self.dom.convert(val) % self.mod

    def modulus(self):
        return self.mod

    def __hash__(self):
        return hash((self.val, self.mod))

    def __repr__(self):
        return "%s(%s)" % (self.__class__.__name__, self.val)

    def __str__(self):
        return "%s mod %s" % (self.val, self.mod)

    def __int__(self):
        return int(self.val)

    def to_int(self):

        sympy_deprecation_warning(
            """ModularInteger.to_int() is deprecated.

            Use int(a) or K = GF(p) and K.to_int(a) instead of a.to_int().
            """,
            deprecated_since_version="1.13",
            active_deprecations_target="modularinteger-to-int",
        )

        if self.sym:
            if self.val <= self.mod // 2:
                return self.val
            else:
                return self.val - self.mod
        else:
            return self.val

    def __pos__(self):
        return self

    def __neg__(self):
        return self.__class__(-self.val)

    @classmethod
    def _get_val(cls, other):
        if isinstance(other, cls):
            return other.val
        else:
            try:
                return cls.dom.convert(other)
            except CoercionFailed:
                return None

    def __add__(self, other):
        val = self._get_val(other)

        if val is not None:
            return self.__class__(self.val + val)
        else:
            return NotImplemented

    def __radd__(self, other):
        return self.__add__(other)

    def __sub__(self, other):
        val = self._get_val(other)

        if val is not None:
            return self.__class__(self.val - val)
        else:
            return NotImplemented

    def __rsub__(self, other):
        return (-self).__add__(other)

    def __mul__(self, other):
        val = self._get_val(other)

        if val is not None:
            return self.__class__(self.val * val)
        else:
            return NotImplemented

    def __rmul__(self, other):
        return self.__mul__(other)

    def __truediv__(self, other):
        val = self._get_val(other)

        if val is not None:
            return self.__class__(self.val * self._invert(val))
        else:
            return NotImplemented

    def __rtruediv__(self, other):
        return self.invert().__mul__(other)

    def __mod__(self, other):
        val = self._get_val(other)

        if val is not None:
            return self.__class__(self.val % val)
        else:
            return NotImplemented

    def __rmod__(self, other):
        val = self._get_val(other)

        if val is not None:
            return self.__class__(val % self.val)
        else:
            return NotImplemented

    def __pow__(self, exp):
        if not exp:
            return self.__class__(self.dom.one)

        if exp < 0:
            val, exp = self.invert().val, -exp
        else:
            val = self.val

        return self.__class__(pow(val, int(exp), self.mod))

    def _compare(self, other, op):
        val = self._get_val(other)

        if val is None:
            return NotImplemented

        return op(self.val, val % self.mod)

    def _compare_deprecated(self, other, op):
        val = self._get_val(other)

        if val is None:
            return NotImplemented

        sympy_deprecation_warning(
            """Ordered comparisons with modular integers are deprecated.

            Use e.g. int(a) < int(b) instead of a < b.
            """,
            deprecated_since_version="1.13",
            active_deprecations_target="modularinteger-compare",
            stacklevel=4,
        )

        return op(self.val, val % self.mod)

    def __eq__(self, other):
        return self._compare(other, operator.eq)

    def __ne__(self, other):
        return self._compare(other, operator.ne)

    def __lt__(self, other):
        return self._compare_deprecated(other, operator.lt)

    def __le__(self, other):
        return self._compare_deprecated(other, operator.le)

    def __gt__(self, other):
        return self._compare_deprecated(other, operator.gt)

    def __ge__(self, other):
        return self._compare_deprecated(other, operator.ge)

    def __bool__(self):
        return bool(self.val)

    @classmethod
    def _invert(cls, value):
        return cls.dom.invert(value, cls.mod)

    def invert(self):
        return self.__class__(self._invert(self.val))

_modular_integer_cache: dict[tuple[Any, Any, Any], type[ModularInteger]] = {}

def ModularIntegerFactory(_mod, _dom, _sym, parent):
    """Create custom class for specific integer modulus."""
    try:
        _mod = _dom.convert(_mod)
    except CoercionFailed:
        ok = False
    else:
        ok = True

    if not ok or _mod < 1:
        raise ValueError("modulus must be a positive integer, got %s" % _mod)

    key = _mod, _dom, _sym

    try:
        cls = _modular_integer_cache[key]
    except KeyError:
        class cls(ModularInteger):
            mod, dom, sym = _mod, _dom, _sym
            _parent = parent

        if _sym:
            cls.__name__ = "SymmetricModularIntegerMod%s" % _mod
        else:
            cls.__name__ = "ModularIntegerMod%s" % _mod

        _modular_integer_cache[key] = cls

    return cls

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\sam2\image_encoder.py ===
# -------------------------------------------------------------------------
# Copyright (R) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
import logging
import warnings

import torch
from sam2.modeling.sam2_base import SAM2Base
from sam2_utils import compare_tensors_with_tolerance, random_sam2_input_image
from torch import nn

import onnxruntime

logger = logging.getLogger(__name__)


class SAM2ImageEncoder(nn.Module):
    def __init__(self, sam_model: SAM2Base) -> None:
        super().__init__()
        self.model = sam_model
        self.image_encoder = sam_model.image_encoder
        self.no_mem_embed = sam_model.no_mem_embed

    def forward(
        self,
        image: torch.Tensor,
        enable_nvtx_profile: bool = False,
    ) -> tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
        """
        Encodes images into features.

        Only supports H=W=1024. If you want to use different image sizes like 512x512,
        see https://github.com/facebookresearch/segment-anything-2/issues/138.

        Args:
            image (torch.Tensor): images of shape [B, 3, H, W], B is batch size, H and W are height and width.
            enable_nvtx_profile (bool): enable NVTX profiling.

        Returns:
            image_features_0: image features of shape [B, 32, H/4, W/4] - high resolution features of level 0
            image_features_1: image features of shape [B, 64, H/8, W/8] - high resolution features of level 1
            image_embeddings: image features of shape [B, 256, H/16, W/16] - 16 is the backbone_stride
        """
        nvtx_helper = None
        if enable_nvtx_profile:
            from nvtx_helper import NvtxHelper

            nvtx_helper = NvtxHelper(["image_encoder", "post_process"])

        if nvtx_helper is not None:
            nvtx_helper.start_profile("image_encoder")

        backbone_out = self.image_encoder(image)

        if nvtx_helper is not None:
            nvtx_helper.stop_profile("image_encoder")
            nvtx_helper.start_profile("post_process")

        # precompute projected level 0 and level 1 features in SAM decoder
        # to avoid running it again on every SAM click
        backbone_out["backbone_fpn"][0] = self.model.sam_mask_decoder.conv_s0(backbone_out["backbone_fpn"][0])
        backbone_out["backbone_fpn"][1] = self.model.sam_mask_decoder.conv_s1(backbone_out["backbone_fpn"][1])

        # Prepare and flatten visual features.
        feature_maps = backbone_out["backbone_fpn"][-self.model.num_feature_levels :]
        vision_pos_embeds = backbone_out["vision_pos_enc"][-self.model.num_feature_levels :]
        feat_sizes = [(x.shape[-2], x.shape[-1]) for x in vision_pos_embeds]

        # flatten NxCxHxW to HWxNxC
        # TODO: we should avoid this transpose since it will be transposed back to NCHW later.
        vision_feats = [x.flatten(2).permute(2, 0, 1) for x in feature_maps]

        vision_feats[-1] = vision_feats[-1] + self.no_mem_embed

        feats = [
            feat.permute(1, 2, 0).reshape(1, -1, *feat_size)
            for feat, feat_size in zip(vision_feats[::-1], feat_sizes[::-1], strict=False)
        ][::-1]

        if nvtx_helper is not None:
            nvtx_helper.stop_profile("post_process")
            nvtx_helper.print_latency()

        return feats[0], feats[1], feats[2]


def export_image_encoder_onnx(
    sam2_model: SAM2Base,
    onnx_model_path: str,
    dynamic_batch_axes: bool = False,
    verbose: bool = False,
    dynamo: bool = False,
    clear_dynamo_metadata: bool = False,
):
    image = random_sam2_input_image()

    sam2_encoder = SAM2ImageEncoder(sam2_model).cpu()
    image_features_0, image_features_1, image_embeddings = sam2_encoder(image)
    logger.info("image.shape: %s", image.shape)
    logger.info("image_features_0.shape: %s", image_features_0.shape)
    logger.info("image_features_1.shape: %s", image_features_1.shape)
    logger.info("image_embeddings.shape: %s", image_embeddings.shape)

    dynamic_axes = None
    if dynamic_batch_axes:
        dynamic_axes = {
            "image": {0: "batch_size"},
            "image_features_0": {0: "batch_size"},
            "image_features_1": {0: "batch_size"},
            "image_embeddings": {0: "batch_size"},
        }

    with warnings.catch_warnings():
        if not verbose:
            warnings.filterwarnings("ignore", category=torch.jit.TracerWarning)
            warnings.filterwarnings("ignore", category=UserWarning)

        if not dynamo:
            torch.onnx.export(
                sam2_encoder,
                image,
                onnx_model_path,
                export_params=True,
                opset_version=17,
                do_constant_folding=True,
                input_names=["image"],
                output_names=["image_features_0", "image_features_1", "image_embeddings"],
                dynamic_axes=dynamic_axes,
            )
        else:
            torch._dynamo.config.capture_scalar_outputs = True
            ep = torch.export.export(
                sam2_encoder,
                args=(image,),
                strict=False,
                dynamic_shapes=[
                    {0: torch.export.Dim.AUTO},
                ],
            )

            onnx_program = torch.onnx.export(
                ep,
                (),
                opset_version=17,
                input_names=["image"],
                output_names=["image_features_0", "image_features_1", "image_embeddings"],
                dynamo=True,
            )
            onnx_program.optimize()
            onnx_program.save(onnx_model_path + ".dynamo.onnx", external_data=False)
            import onnx

            from onnxruntime.transformers.dynamo_onnx_helper import DynamoOnnxHelper

            onnx_model = onnx.load_model(onnx_model_path + ".dynamo.onnx", load_external_data=True)
            if dynamic_batch_axes:
                # Fix labels of dynamic axes since they can't be specified during Dynamo export currently
                onnx_model.graph.input[0].type.tensor_type.shape.dim[0].dim_param = "batch_size"
                for i in range(3):
                    onnx_model.graph.output[i].type.tensor_type.shape.dim[0].dim_param = "batch_size"

            onnx_model_helper = DynamoOnnxHelper(onnx_model)
            onnx_model_helper.convert_constants_to_initializers()
            if clear_dynamo_metadata:
                onnx_model_helper.clear_metadata()

            import os

            if os.path.exists(onnx_model_path):
                os.remove(onnx_model_path)
            if os.path.exists(onnx_model_path + ".data"):
                os.remove(onnx_model_path + ".data")
            onnx_model_helper.model.save_model_to_file(
                onnx_model_path, use_external_data_format=True, all_tensors_to_one_file=True, convert_attribute=True
            )

    print("encoder onnx model saved to", onnx_model_path)


def test_image_encoder_onnx(
    sam2_model: SAM2Base,
    onnx_model_path: str,
    dynamic_batch_axes=False,
):
    ort_session = onnxruntime.InferenceSession(onnx_model_path, providers=["CPUExecutionProvider"])

    model_inputs = ort_session.get_inputs()
    input_names = [model_inputs[i].name for i in range(len(model_inputs))]
    logger.info("input_names: %s", input_names)

    model_outputs = ort_session.get_outputs()
    output_names = [model_outputs[i].name for i in range(len(model_outputs))]
    logger.info("output_names: %s", output_names)

    batch_sizes = [1, 2] if dynamic_batch_axes else [1]
    for batch_size in batch_sizes:
        image = random_sam2_input_image(batch_size)

        sam2_encoder = SAM2ImageEncoder(sam2_model).cpu()
        image_features_0, image_features_1, image_embeddings = sam2_encoder(image.clone())

        logger.info("image.shape: %s", image.shape)
        logger.info("image_features_0.shape: %s", image_features_0.shape)
        logger.info("image_features_1.shape: %s", image_features_1.shape)
        logger.info("image_embeddings.shape: %s", image_embeddings.shape)

        outputs = ort_session.run(output_names, {"image": image.numpy()})
        for i, output_name in enumerate(output_names):
            logger.info("output %s shape %s", output_name, outputs[i].shape)
        ort_image_features_0, ort_image_features_1, ort_image_embeddings = outputs

        # ONNXRuntime and PyTorch has about 0.75% mismatched elements, but seems not impacting segmentation results.
        if (
            compare_tensors_with_tolerance(
                "image_features_0",
                image_features_0,
                torch.tensor(ort_image_features_0),
                mismatch_percentage_tolerance=1,
            )
            and compare_tensors_with_tolerance(
                "image_features_1",
                image_features_1,
                torch.tensor(ort_image_features_1),
                mismatch_percentage_tolerance=1,
            )
            and compare_tensors_with_tolerance(
                "image_embeddings",
                image_embeddings,
                torch.tensor(ort_image_embeddings),
                mismatch_percentage_tolerance=1,
            )
        ):
            print(f"onnx model has been verified for batch_size={batch_size}: {onnx_model_path}")
        else:
            print(f"onnx model verification failed for batch_size={batch_size}: {onnx_model_path}")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\werkzeug\middleware\http_proxy.py ===
"""
Basic HTTP Proxy
================

.. autoclass:: ProxyMiddleware

:copyright: 2007 Pallets
:license: BSD-3-Clause
"""

from __future__ import annotations

import typing as t
from http import client
from urllib.parse import quote
from urllib.parse import urlsplit

from ..datastructures import EnvironHeaders
from ..http import is_hop_by_hop_header
from ..wsgi import get_input_stream

if t.TYPE_CHECKING:
    from _typeshed.wsgi import StartResponse
    from _typeshed.wsgi import WSGIApplication
    from _typeshed.wsgi import WSGIEnvironment


class ProxyMiddleware:
    """Proxy requests under a path to an external server, routing other
    requests to the app.

    This middleware can only proxy HTTP requests, as HTTP is the only
    protocol handled by the WSGI server. Other protocols, such as
    WebSocket requests, cannot be proxied at this layer. This should
    only be used for development, in production a real proxy server
    should be used.

    The middleware takes a dict mapping a path prefix to a dict
    describing the host to be proxied to::

        app = ProxyMiddleware(app, {
            "/static/": {
                "target": "http://127.0.0.1:5001/",
            }
        })

    Each host has the following options:

    ``target``:
        The target URL to dispatch to. This is required.
    ``remove_prefix``:
        Whether to remove the prefix from the URL before dispatching it
        to the target. The default is ``False``.
    ``host``:
        ``"<auto>"`` (default):
            The host header is automatically rewritten to the URL of the
            target.
        ``None``:
            The host header is unmodified from the client request.
        Any other value:
            The host header is overwritten with the value.
    ``headers``:
        A dictionary of headers to be sent with the request to the
        target. The default is ``{}``.
    ``ssl_context``:
        A :class:`ssl.SSLContext` defining how to verify requests if the
        target is HTTPS. The default is ``None``.

    In the example above, everything under ``"/static/"`` is proxied to
    the server on port 5001. The host header is rewritten to the target,
    and the ``"/static/"`` prefix is removed from the URLs.

    :param app: The WSGI application to wrap.
    :param targets: Proxy target configurations. See description above.
    :param chunk_size: Size of chunks to read from input stream and
        write to target.
    :param timeout: Seconds before an operation to a target fails.

    .. versionadded:: 0.14
    """

    def __init__(
        self,
        app: WSGIApplication,
        targets: t.Mapping[str, dict[str, t.Any]],
        chunk_size: int = 2 << 13,
        timeout: int = 10,
    ) -> None:
        def _set_defaults(opts: dict[str, t.Any]) -> dict[str, t.Any]:
            opts.setdefault("remove_prefix", False)
            opts.setdefault("host", "<auto>")
            opts.setdefault("headers", {})
            opts.setdefault("ssl_context", None)
            return opts

        self.app = app
        self.targets = {
            f"/{k.strip('/')}/": _set_defaults(v) for k, v in targets.items()
        }
        self.chunk_size = chunk_size
        self.timeout = timeout

    def proxy_to(
        self, opts: dict[str, t.Any], path: str, prefix: str
    ) -> WSGIApplication:
        target = urlsplit(opts["target"])
        # socket can handle unicode host, but header must be ascii
        host = target.hostname.encode("idna").decode("ascii")

        def application(
            environ: WSGIEnvironment, start_response: StartResponse
        ) -> t.Iterable[bytes]:
            headers = list(EnvironHeaders(environ).items())
            headers[:] = [
                (k, v)
                for k, v in headers
                if not is_hop_by_hop_header(k)
                and k.lower() not in ("content-length", "host")
            ]
            headers.append(("Connection", "close"))

            if opts["host"] == "<auto>":
                headers.append(("Host", host))
            elif opts["host"] is None:
                headers.append(("Host", environ["HTTP_HOST"]))
            else:
                headers.append(("Host", opts["host"]))

            headers.extend(opts["headers"].items())
            remote_path = path

            if opts["remove_prefix"]:
                remote_path = remote_path[len(prefix) :].lstrip("/")
                remote_path = f"{target.path.rstrip('/')}/{remote_path}"

            content_length = environ.get("CONTENT_LENGTH")
            chunked = False

            if content_length not in ("", None):
                headers.append(("Content-Length", content_length))  # type: ignore
            elif content_length is not None:
                headers.append(("Transfer-Encoding", "chunked"))
                chunked = True

            try:
                if target.scheme == "http":
                    con = client.HTTPConnection(
                        host, target.port or 80, timeout=self.timeout
                    )
                elif target.scheme == "https":
                    con = client.HTTPSConnection(
                        host,
                        target.port or 443,
                        timeout=self.timeout,
                        context=opts["ssl_context"],
                    )
                else:
                    raise RuntimeError(
                        "Target scheme must be 'http' or 'https', got"
                        f" {target.scheme!r}."
                    )

                con.connect()
                # safe = https://url.spec.whatwg.org/#url-path-segment-string
                # as well as percent for things that are already quoted
                remote_url = quote(remote_path, safe="!$&'()*+,/:;=@%")
                querystring = environ["QUERY_STRING"]

                if querystring:
                    remote_url = f"{remote_url}?{querystring}"

                con.putrequest(environ["REQUEST_METHOD"], remote_url, skip_host=True)

                for k, v in headers:
                    if k.lower() == "connection":
                        v = "close"

                    con.putheader(k, v)

                con.endheaders()
                stream = get_input_stream(environ)

                while True:
                    data = stream.read(self.chunk_size)

                    if not data:
                        break

                    if chunked:
                        con.send(b"%x\r\n%s\r\n" % (len(data), data))
                    else:
                        con.send(data)

                resp = con.getresponse()
            except OSError:
                from ..exceptions import BadGateway

                return BadGateway()(environ, start_response)

            start_response(
                f"{resp.status} {resp.reason}",
                [
                    (k.title(), v)
                    for k, v in resp.getheaders()
                    if not is_hop_by_hop_header(k)
                ],
            )

            def read() -> t.Iterator[bytes]:
                while True:
                    try:
                        data = resp.read(self.chunk_size)
                    except OSError:
                        break

                    if not data:
                        break

                    yield data

            return read()

        return application

    def __call__(
        self, environ: WSGIEnvironment, start_response: StartResponse
    ) -> t.Iterable[bytes]:
        path = environ["PATH_INFO"]
        app = self.app

        for prefix, opts in self.targets.items():
            if path.startswith(prefix):
                app = self.proxy_to(opts, path, prefix)
                break

        return app(environ, start_response)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\fft\_helper.py ===
"""
Discrete Fourier Transforms - _helper.py

"""
from numpy._core import arange, asarray, empty, integer, roll
from numpy._core.overrides import array_function_dispatch, set_module

# Created by Pearu Peterson, September 2002

__all__ = ['fftshift', 'ifftshift', 'fftfreq', 'rfftfreq']

integer_types = (int, integer)


def _fftshift_dispatcher(x, axes=None):
    return (x,)


@array_function_dispatch(_fftshift_dispatcher, module='numpy.fft')
def fftshift(x, axes=None):
    """
    Shift the zero-frequency component to the center of the spectrum.

    This function swaps half-spaces for all axes listed (defaults to all).
    Note that ``y[0]`` is the Nyquist component only if ``len(x)`` is even.

    Parameters
    ----------
    x : array_like
        Input array.
    axes : int or shape tuple, optional
        Axes over which to shift.  Default is None, which shifts all axes.

    Returns
    -------
    y : ndarray
        The shifted array.

    See Also
    --------
    ifftshift : The inverse of `fftshift`.

    Examples
    --------
    >>> import numpy as np
    >>> freqs = np.fft.fftfreq(10, 0.1)
    >>> freqs
    array([ 0.,  1.,  2., ..., -3., -2., -1.])
    >>> np.fft.fftshift(freqs)
    array([-5., -4., -3., -2., -1.,  0.,  1.,  2.,  3.,  4.])

    Shift the zero-frequency component only along the second axis:

    >>> freqs = np.fft.fftfreq(9, d=1./9).reshape(3, 3)
    >>> freqs
    array([[ 0.,  1.,  2.],
           [ 3.,  4., -4.],
           [-3., -2., -1.]])
    >>> np.fft.fftshift(freqs, axes=(1,))
    array([[ 2.,  0.,  1.],
           [-4.,  3.,  4.],
           [-1., -3., -2.]])

    """
    x = asarray(x)
    if axes is None:
        axes = tuple(range(x.ndim))
        shift = [dim // 2 for dim in x.shape]
    elif isinstance(axes, integer_types):
        shift = x.shape[axes] // 2
    else:
        shift = [x.shape[ax] // 2 for ax in axes]

    return roll(x, shift, axes)


@array_function_dispatch(_fftshift_dispatcher, module='numpy.fft')
def ifftshift(x, axes=None):
    """
    The inverse of `fftshift`. Although identical for even-length `x`, the
    functions differ by one sample for odd-length `x`.

    Parameters
    ----------
    x : array_like
        Input array.
    axes : int or shape tuple, optional
        Axes over which to calculate.  Defaults to None, which shifts all axes.

    Returns
    -------
    y : ndarray
        The shifted array.

    See Also
    --------
    fftshift : Shift zero-frequency component to the center of the spectrum.

    Examples
    --------
    >>> import numpy as np
    >>> freqs = np.fft.fftfreq(9, d=1./9).reshape(3, 3)
    >>> freqs
    array([[ 0.,  1.,  2.],
           [ 3.,  4., -4.],
           [-3., -2., -1.]])
    >>> np.fft.ifftshift(np.fft.fftshift(freqs))
    array([[ 0.,  1.,  2.],
           [ 3.,  4., -4.],
           [-3., -2., -1.]])

    """
    x = asarray(x)
    if axes is None:
        axes = tuple(range(x.ndim))
        shift = [-(dim // 2) for dim in x.shape]
    elif isinstance(axes, integer_types):
        shift = -(x.shape[axes] // 2)
    else:
        shift = [-(x.shape[ax] // 2) for ax in axes]

    return roll(x, shift, axes)


@set_module('numpy.fft')
def fftfreq(n, d=1.0, device=None):
    """
    Return the Discrete Fourier Transform sample frequencies.

    The returned float array `f` contains the frequency bin centers in cycles
    per unit of the sample spacing (with zero at the start).  For instance, if
    the sample spacing is in seconds, then the frequency unit is cycles/second.

    Given a window length `n` and a sample spacing `d`::

      f = [0, 1, ...,   n/2-1,     -n/2, ..., -1] / (d*n)   if n is even
      f = [0, 1, ..., (n-1)/2, -(n-1)/2, ..., -1] / (d*n)   if n is odd

    Parameters
    ----------
    n : int
        Window length.
    d : scalar, optional
        Sample spacing (inverse of the sampling rate). Defaults to 1.
    device : str, optional
        The device on which to place the created array. Default: ``None``.
        For Array-API interoperability only, so must be ``"cpu"`` if passed.

        .. versionadded:: 2.0.0

    Returns
    -------
    f : ndarray
        Array of length `n` containing the sample frequencies.

    Examples
    --------
    >>> import numpy as np
    >>> signal = np.array([-2, 8, 6, 4, 1, 0, 3, 5], dtype=float)
    >>> fourier = np.fft.fft(signal)
    >>> n = signal.size
    >>> timestep = 0.1
    >>> freq = np.fft.fftfreq(n, d=timestep)
    >>> freq
    array([ 0.  ,  1.25,  2.5 , ..., -3.75, -2.5 , -1.25])

    """
    if not isinstance(n, integer_types):
        raise ValueError("n should be an integer")
    val = 1.0 / (n * d)
    results = empty(n, int, device=device)
    N = (n - 1) // 2 + 1
    p1 = arange(0, N, dtype=int, device=device)
    results[:N] = p1
    p2 = arange(-(n // 2), 0, dtype=int, device=device)
    results[N:] = p2
    return results * val


@set_module('numpy.fft')
def rfftfreq(n, d=1.0, device=None):
    """
    Return the Discrete Fourier Transform sample frequencies
    (for usage with rfft, irfft).

    The returned float array `f` contains the frequency bin centers in cycles
    per unit of the sample spacing (with zero at the start).  For instance, if
    the sample spacing is in seconds, then the frequency unit is cycles/second.

    Given a window length `n` and a sample spacing `d`::

      f = [0, 1, ...,     n/2-1,     n/2] / (d*n)   if n is even
      f = [0, 1, ..., (n-1)/2-1, (n-1)/2] / (d*n)   if n is odd

    Unlike `fftfreq` (but like `scipy.fftpack.rfftfreq`)
    the Nyquist frequency component is considered to be positive.

    Parameters
    ----------
    n : int
        Window length.
    d : scalar, optional
        Sample spacing (inverse of the sampling rate). Defaults to 1.
    device : str, optional
        The device on which to place the created array. Default: ``None``.
        For Array-API interoperability only, so must be ``"cpu"`` if passed.

        .. versionadded:: 2.0.0

    Returns
    -------
    f : ndarray
        Array of length ``n//2 + 1`` containing the sample frequencies.

    Examples
    --------
    >>> import numpy as np
    >>> signal = np.array([-2, 8, 6, 4, 1, 0, 3, 5, -3, 4], dtype=float)
    >>> fourier = np.fft.rfft(signal)
    >>> n = signal.size
    >>> sample_rate = 100
    >>> freq = np.fft.fftfreq(n, d=1./sample_rate)
    >>> freq
    array([  0.,  10.,  20., ..., -30., -20., -10.])
    >>> freq = np.fft.rfftfreq(n, d=1./sample_rate)
    >>> freq
    array([  0.,  10.,  20.,  30.,  40.,  50.])

    """
    if not isinstance(n, integer_types):
        raise ValueError("n should be an integer")
    val = 1.0 / (n * d)
    N = n // 2 + 1
    results = arange(0, N, dtype=int, device=device)
    return results * val

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\functions\special\singularity_functions.py ===
from sympy.core import S, oo, diff
from sympy.core.function import DefinedFunction, ArgumentIndexError
from sympy.core.logic import fuzzy_not
from sympy.core.relational import Eq
from sympy.functions.elementary.complexes import im
from sympy.functions.elementary.piecewise import Piecewise
from sympy.functions.special.delta_functions import Heaviside

###############################################################################
############################# SINGULARITY FUNCTION ############################
###############################################################################


class SingularityFunction(DefinedFunction):
    r"""
    Singularity functions are a class of discontinuous functions.

    Explanation
    ===========

    Singularity functions take a variable, an offset, and an exponent as
    arguments. These functions are represented using Macaulay brackets as:

    SingularityFunction(x, a, n) := <x - a>^n

    The singularity function will automatically evaluate to
    ``Derivative(DiracDelta(x - a), x, -n - 1)`` if ``n < 0``
    and ``(x - a)**n*Heaviside(x - a, 1)`` if ``n >= 0``.

    Examples
    ========

    >>> from sympy import SingularityFunction, diff, Piecewise, DiracDelta, Heaviside, Symbol
    >>> from sympy.abc import x, a, n
    >>> SingularityFunction(x, a, n)
    SingularityFunction(x, a, n)
    >>> y = Symbol('y', positive=True)
    >>> n = Symbol('n', nonnegative=True)
    >>> SingularityFunction(y, -10, n)
    (y + 10)**n
    >>> y = Symbol('y', negative=True)
    >>> SingularityFunction(y, 10, n)
    0
    >>> SingularityFunction(x, 4, -1).subs(x, 4)
    oo
    >>> SingularityFunction(x, 10, -2).subs(x, 10)
    oo
    >>> SingularityFunction(4, 1, 5)
    243
    >>> diff(SingularityFunction(x, 1, 5) + SingularityFunction(x, 1, 4), x)
    4*SingularityFunction(x, 1, 3) + 5*SingularityFunction(x, 1, 4)
    >>> diff(SingularityFunction(x, 4, 0), x, 2)
    SingularityFunction(x, 4, -2)
    >>> SingularityFunction(x, 4, 5).rewrite(Piecewise)
    Piecewise(((x - 4)**5, x >= 4), (0, True))
    >>> expr = SingularityFunction(x, a, n)
    >>> y = Symbol('y', positive=True)
    >>> n = Symbol('n', nonnegative=True)
    >>> expr.subs({x: y, a: -10, n: n})
    (y + 10)**n

    The methods ``rewrite(DiracDelta)``, ``rewrite(Heaviside)``, and
    ``rewrite('HeavisideDiracDelta')`` returns the same output. One can use any
    of these methods according to their choice.

    >>> expr = SingularityFunction(x, 4, 5) + SingularityFunction(x, -3, -1) - SingularityFunction(x, 0, -2)
    >>> expr.rewrite(Heaviside)
    (x - 4)**5*Heaviside(x - 4, 1) + DiracDelta(x + 3) - DiracDelta(x, 1)
    >>> expr.rewrite(DiracDelta)
    (x - 4)**5*Heaviside(x - 4, 1) + DiracDelta(x + 3) - DiracDelta(x, 1)
    >>> expr.rewrite('HeavisideDiracDelta')
    (x - 4)**5*Heaviside(x - 4, 1) + DiracDelta(x + 3) - DiracDelta(x, 1)

    See Also
    ========

    DiracDelta, Heaviside

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Singularity_function

    """

    is_real = True

    def fdiff(self, argindex=1):
        """
        Returns the first derivative of a DiracDelta Function.

        Explanation
        ===========

        The difference between ``diff()`` and ``fdiff()`` is: ``diff()`` is the
        user-level function and ``fdiff()`` is an object method. ``fdiff()`` is
        a convenience method available in the ``Function`` class. It returns
        the derivative of the function without considering the chain rule.
        ``diff(function, x)`` calls ``Function._eval_derivative`` which in turn
        calls ``fdiff()`` internally to compute the derivative of the function.

        """

        if argindex == 1:
            x, a, n = self.args
            if n in (S.Zero, S.NegativeOne, S(-2), S(-3)):
                return self.func(x, a, n-1)
            elif n.is_positive:
                return n*self.func(x, a, n-1)
        else:
            raise ArgumentIndexError(self, argindex)

    @classmethod
    def eval(cls, variable, offset, exponent):
        """
        Returns a simplified form or a value of Singularity Function depending
        on the argument passed by the object.

        Explanation
        ===========

        The ``eval()`` method is automatically called when the
        ``SingularityFunction`` class is about to be instantiated and it
        returns either some simplified instance or the unevaluated instance
        depending on the argument passed. In other words, ``eval()`` method is
        not needed to be called explicitly, it is being called and evaluated
        once the object is called.

        Examples
        ========

        >>> from sympy import SingularityFunction, Symbol, nan
        >>> from sympy.abc import x, a, n
        >>> SingularityFunction(x, a, n)
        SingularityFunction(x, a, n)
        >>> SingularityFunction(5, 3, 2)
        4
        >>> SingularityFunction(x, a, nan)
        nan
        >>> SingularityFunction(x, 3, 0).subs(x, 3)
        1
        >>> SingularityFunction(4, 1, 5)
        243
        >>> x = Symbol('x', positive = True)
        >>> a = Symbol('a', negative = True)
        >>> n = Symbol('n', nonnegative = True)
        >>> SingularityFunction(x, a, n)
        (-a + x)**n
        >>> x = Symbol('x', negative = True)
        >>> a = Symbol('a', positive = True)
        >>> SingularityFunction(x, a, n)
        0

        """

        x = variable
        a = offset
        n = exponent
        shift = (x - a)

        if fuzzy_not(im(shift).is_zero):
            raise ValueError("Singularity Functions are defined only for Real Numbers.")
        if fuzzy_not(im(n).is_zero):
            raise ValueError("Singularity Functions are not defined for imaginary exponents.")
        if shift is S.NaN or n is S.NaN:
            return S.NaN
        if (n + 4).is_negative:
            raise ValueError("Singularity Functions are not defined for exponents less than -4.")
        if shift.is_extended_negative:
            return S.Zero
        if n.is_nonnegative:
            if shift.is_zero:  # use literal 0 in case of Symbol('z', zero=True)
                return S.Zero**n
            if shift.is_extended_nonnegative:
                return shift**n
        if n in (S.NegativeOne, -2, -3, -4):
            if shift.is_negative or shift.is_extended_positive:
                return S.Zero
            if shift.is_zero:
                return oo

    def _eval_rewrite_as_Piecewise(self, *args, **kwargs):
        '''
        Converts a Singularity Function expression into its Piecewise form.

        '''
        x, a, n = self.args

        if n in (S.NegativeOne, S(-2), S(-3), S(-4)):
            return Piecewise((oo, Eq(x - a, 0)), (0, True))
        elif n.is_nonnegative:
            return Piecewise(((x - a)**n, x - a >= 0), (0, True))

    def _eval_rewrite_as_Heaviside(self, *args, **kwargs):
        '''
        Rewrites a Singularity Function expression using Heavisides and DiracDeltas.

        '''
        x, a, n = self.args

        if n == -4:
            return diff(Heaviside(x - a), x.free_symbols.pop(), 4)
        if n == -3:
            return diff(Heaviside(x - a), x.free_symbols.pop(), 3)
        if n == -2:
            return diff(Heaviside(x - a), x.free_symbols.pop(), 2)
        if n == -1:
            return diff(Heaviside(x - a), x.free_symbols.pop(), 1)
        if n.is_nonnegative:
            return (x - a)**n*Heaviside(x - a, 1)

    def _eval_as_leading_term(self, x, logx, cdir):
        z, a, n = self.args
        shift = (z - a).subs(x, 0)
        if n < 0:
            return S.Zero
        elif n.is_zero and shift.is_zero:
            return S.Zero if cdir == -1 else S.One
        elif shift.is_positive:
            return shift**n
        return S.Zero

    def _eval_nseries(self, x, n, logx=None, cdir=0):
        z, a, n = self.args
        shift = (z - a).subs(x, 0)
        if n < 0:
            return S.Zero
        elif n.is_zero and shift.is_zero:
            return S.Zero if cdir == -1 else S.One
        elif shift.is_positive:
            return ((z - a)**n)._eval_nseries(x, n, logx=logx, cdir=cdir)
        return S.Zero

    _eval_rewrite_as_DiracDelta = _eval_rewrite_as_Heaviside
    _eval_rewrite_as_HeavisideDiracDelta = _eval_rewrite_as_Heaviside

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\support\event_firing_webdriver.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

from typing import Any, List, Tuple

from selenium.common.exceptions import WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.remote.webelement import WebElement

from .abstract_event_listener import AbstractEventListener


def _wrap_elements(result, ef_driver):
    # handle the case if another wrapper wraps EventFiringWebElement
    if isinstance(result, EventFiringWebElement):
        return result
    if isinstance(result, WebElement):
        return EventFiringWebElement(result, ef_driver)
    if isinstance(result, list):
        return [_wrap_elements(item, ef_driver) for item in result]
    return result


class EventFiringWebDriver:
    """A wrapper around an arbitrary WebDriver instance which supports firing
    events."""

    def __init__(self, driver: WebDriver, event_listener: AbstractEventListener) -> None:
        """Creates a new instance of the EventFiringWebDriver.

        :Args:
         - driver : A WebDriver instance
         - event_listener : Instance of a class that subclasses AbstractEventListener and implements it fully
                            or partially

        Example:

        ::

            from selenium.webdriver import Firefox
            from selenium.webdriver.support.events import EventFiringWebDriver, AbstractEventListener


            class MyListener(AbstractEventListener):
                def before_navigate_to(self, url, driver):
                    print("Before navigate to %s" % url)

                def after_navigate_to(self, url, driver):
                    print("After navigate to %s" % url)


            driver = Firefox()
            ef_driver = EventFiringWebDriver(driver, MyListener())
            ef_driver.get("http://www.google.co.in/")
        """
        if not isinstance(driver, WebDriver):
            raise WebDriverException("A WebDriver instance must be supplied")
        if not isinstance(event_listener, AbstractEventListener):
            raise WebDriverException("Event listener must be a subclass of AbstractEventListener")
        self._driver = driver
        self._driver._wrap_value = self._wrap_value
        self._listener = event_listener

    @property
    def wrapped_driver(self) -> WebDriver:
        """Returns the WebDriver instance wrapped by this
        EventsFiringWebDriver."""
        return self._driver

    def get(self, url: str) -> None:
        self._dispatch("navigate_to", (url, self._driver), "get", (url,))

    def back(self) -> None:
        self._dispatch("navigate_back", (self._driver,), "back", ())

    def forward(self) -> None:
        self._dispatch("navigate_forward", (self._driver,), "forward", ())

    def execute_script(self, script: str, *args):
        unwrapped_args = (script,) + self._unwrap_element_args(args)
        return self._dispatch("execute_script", (script, self._driver), "execute_script", unwrapped_args)

    def execute_async_script(self, script, *args):
        unwrapped_args = (script,) + self._unwrap_element_args(args)
        return self._dispatch("execute_script", (script, self._driver), "execute_async_script", unwrapped_args)

    def close(self) -> None:
        self._dispatch("close", (self._driver,), "close", ())

    def quit(self) -> None:
        self._dispatch("quit", (self._driver,), "quit", ())

    def find_element(self, by=By.ID, value=None) -> WebElement:
        return self._dispatch("find", (by, value, self._driver), "find_element", (by, value))

    def find_elements(self, by=By.ID, value=None) -> List[WebElement]:
        return self._dispatch("find", (by, value, self._driver), "find_elements", (by, value))

    def _dispatch(self, l_call: str, l_args: Tuple[Any, ...], d_call: str, d_args: Tuple[Any, ...]):
        getattr(self._listener, f"before_{l_call}")(*l_args)
        try:
            result = getattr(self._driver, d_call)(*d_args)
        except Exception as exc:
            self._listener.on_exception(exc, self._driver)
            raise
        getattr(self._listener, f"after_{l_call}")(*l_args)
        return _wrap_elements(result, self)

    def _unwrap_element_args(self, args):
        if isinstance(args, EventFiringWebElement):
            return args.wrapped_element
        if isinstance(args, tuple):
            return tuple(self._unwrap_element_args(item) for item in args)
        if isinstance(args, list):
            return [self._unwrap_element_args(item) for item in args]
        return args

    def _wrap_value(self, value):
        if isinstance(value, EventFiringWebElement):
            return WebDriver._wrap_value(self._driver, value.wrapped_element)
        return WebDriver._wrap_value(self._driver, value)

    def __setattr__(self, item, value):
        if item.startswith("_") or not hasattr(self._driver, item):
            object.__setattr__(self, item, value)
        else:
            try:
                object.__setattr__(self._driver, item, value)
            except Exception as exc:
                self._listener.on_exception(exc, self._driver)
                raise

    def __getattr__(self, name):
        def _wrap(*args, **kwargs):
            try:
                result = attrib(*args, **kwargs)
                return _wrap_elements(result, self)
            except Exception as exc:
                self._listener.on_exception(exc, self._driver)
                raise

        try:
            attrib = getattr(self._driver, name)
            return _wrap if callable(attrib) else attrib
        except Exception as exc:
            self._listener.on_exception(exc, self._driver)
            raise


class EventFiringWebElement:
    """A wrapper around WebElement instance which supports firing events."""

    def __init__(self, webelement: WebElement, ef_driver: EventFiringWebDriver) -> None:
        """Creates a new instance of the EventFiringWebElement."""
        self._webelement = webelement
        self._ef_driver = ef_driver
        self._driver = ef_driver.wrapped_driver
        self._listener = ef_driver._listener

    @property
    def wrapped_element(self) -> WebElement:
        """Returns the WebElement wrapped by this EventFiringWebElement
        instance."""
        return self._webelement

    def click(self) -> None:
        self._dispatch("click", (self._webelement, self._driver), "click", ())

    def clear(self) -> None:
        self._dispatch("change_value_of", (self._webelement, self._driver), "clear", ())

    def send_keys(self, *value) -> None:
        self._dispatch("change_value_of", (self._webelement, self._driver), "send_keys", value)

    def find_element(self, by=By.ID, value=None) -> WebElement:
        return self._dispatch("find", (by, value, self._driver), "find_element", (by, value))

    def find_elements(self, by=By.ID, value=None) -> List[WebElement]:
        return self._dispatch("find", (by, value, self._driver), "find_elements", (by, value))

    def _dispatch(self, l_call, l_args, d_call, d_args):
        getattr(self._listener, f"before_{l_call}")(*l_args)
        try:
            result = getattr(self._webelement, d_call)(*d_args)
        except Exception as exc:
            self._listener.on_exception(exc, self._driver)
            raise
        getattr(self._listener, f"after_{l_call}")(*l_args)
        return _wrap_elements(result, self._ef_driver)

    def __setattr__(self, item, value):
        if item.startswith("_") or not hasattr(self._webelement, item):
            object.__setattr__(self, item, value)
        else:
            try:
                object.__setattr__(self._webelement, item, value)
            except Exception as exc:
                self._listener.on_exception(exc, self._driver)
                raise

    def __getattr__(self, name):
        def _wrap(*args, **kwargs):
            try:
                result = attrib(*args, **kwargs)
                return _wrap_elements(result, self._ef_driver)
            except Exception as exc:
                self._listener.on_exception(exc, self._driver)
                raise

        try:
            attrib = getattr(self._webelement, name)
            return _wrap if callable(attrib) else attrib
        except Exception as exc:
            self._listener.on_exception(exc, self._driver)
            raise


# Register a virtual subclass.
WebElement.register(EventFiringWebElement)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\unify\core.py ===
""" Generic Unification algorithm for expression trees with lists of children

This implementation is a direct translation of

Artificial Intelligence: A Modern Approach by Stuart Russel and Peter Norvig
Second edition, section 9.2, page 276

It is modified in the following ways:

1.  We allow associative and commutative Compound expressions. This results in
    combinatorial blowup.
2.  We explore the tree lazily.
3.  We provide generic interfaces to symbolic algebra libraries in Python.

A more traditional version can be found here
http://aima.cs.berkeley.edu/python/logic.html
"""

from sympy.utilities.iterables import kbins

class Compound:
    """ A little class to represent an interior node in the tree

    This is analogous to SymPy.Basic for non-Atoms
    """
    def __init__(self, op, args):
        self.op = op
        self.args = args

    def __eq__(self, other):
        return (type(self) is type(other) and self.op == other.op and
                self.args == other.args)

    def __hash__(self):
        return hash((type(self), self.op, self.args))

    def __str__(self):
        return "%s[%s]" % (str(self.op), ', '.join(map(str, self.args)))

class Variable:
    """ A Wild token """
    def __init__(self, arg):
        self.arg = arg

    def __eq__(self, other):
        return type(self) is type(other) and self.arg == other.arg

    def __hash__(self):
        return hash((type(self), self.arg))

    def __str__(self):
        return "Variable(%s)" % str(self.arg)

class CondVariable:
    """ A wild token that matches conditionally.

    arg   - a wild token.
    valid - an additional constraining function on a match.
    """
    def __init__(self, arg, valid):
        self.arg = arg
        self.valid = valid

    def __eq__(self, other):
        return (type(self) is type(other) and
                self.arg == other.arg and
                self.valid == other.valid)

    def __hash__(self):
        return hash((type(self), self.arg, self.valid))

    def __str__(self):
        return "CondVariable(%s)" % str(self.arg)

def unify(x, y, s=None, **fns):
    """ Unify two expressions.

    Parameters
    ==========

        x, y - expression trees containing leaves, Compounds and Variables.
        s    - a mapping of variables to subtrees.

    Returns
    =======

        lazy sequence of mappings {Variable: subtree}

    Examples
    ========

    >>> from sympy.unify.core import unify, Compound, Variable
    >>> expr    = Compound("Add", ("x", "y"))
    >>> pattern = Compound("Add", ("x", Variable("a")))
    >>> next(unify(expr, pattern, {}))
    {Variable(a): 'y'}
    """
    s = s or {}

    if x == y:
        yield s
    elif isinstance(x, (Variable, CondVariable)):
        yield from unify_var(x, y, s, **fns)
    elif isinstance(y, (Variable, CondVariable)):
        yield from unify_var(y, x, s, **fns)
    elif isinstance(x, Compound) and isinstance(y, Compound):
        is_commutative = fns.get('is_commutative', lambda x: False)
        is_associative = fns.get('is_associative', lambda x: False)
        for sop in unify(x.op, y.op, s, **fns):
            if is_associative(x) and is_associative(y):
                a, b = (x, y) if len(x.args) < len(y.args) else (y, x)
                if is_commutative(x) and is_commutative(y):
                    combs = allcombinations(a.args, b.args, 'commutative')
                else:
                    combs = allcombinations(a.args, b.args, 'associative')
                for aaargs, bbargs in combs:
                    aa = [unpack(Compound(a.op, arg)) for arg in aaargs]
                    bb = [unpack(Compound(b.op, arg)) for arg in bbargs]
                    yield from unify(aa, bb, sop, **fns)
            elif len(x.args) == len(y.args):
                yield from unify(x.args, y.args, sop, **fns)

    elif is_args(x) and is_args(y) and len(x) == len(y):
        if len(x) == 0:
            yield s
        else:
            for shead in unify(x[0], y[0], s, **fns):
                yield from unify(x[1:], y[1:], shead, **fns)

def unify_var(var, x, s, **fns):
    if var in s:
        yield from unify(s[var], x, s, **fns)
    elif occur_check(var, x):
        pass
    elif isinstance(var, CondVariable) and var.valid(x):
        yield assoc(s, var, x)
    elif isinstance(var, Variable):
        yield assoc(s, var, x)

def occur_check(var, x):
    """ var occurs in subtree owned by x? """
    if var == x:
        return True
    elif isinstance(x, Compound):
        return occur_check(var, x.args)
    elif is_args(x):
        if any(occur_check(var, xi) for xi in x): return True
    return False

def assoc(d, key, val):
    """ Return copy of d with key associated to val """
    d = d.copy()
    d[key] = val
    return d

def is_args(x):
    """ Is x a traditional iterable? """
    return type(x) in (tuple, list, set)

def unpack(x):
    if isinstance(x, Compound) and len(x.args) == 1:
        return x.args[0]
    else:
        return x

def allcombinations(A, B, ordered):
    """
    Restructure A and B to have the same number of elements.

    Parameters
    ==========

    ordered must be either 'commutative' or 'associative'.

    A and B can be rearranged so that the larger of the two lists is
    reorganized into smaller sublists.

    Examples
    ========

    >>> from sympy.unify.core import allcombinations
    >>> for x in allcombinations((1, 2, 3), (5, 6), 'associative'): print(x)
    (((1,), (2, 3)), ((5,), (6,)))
    (((1, 2), (3,)), ((5,), (6,)))

    >>> for x in allcombinations((1, 2, 3), (5, 6), 'commutative'): print(x)
        (((1,), (2, 3)), ((5,), (6,)))
        (((1, 2), (3,)), ((5,), (6,)))
        (((1,), (3, 2)), ((5,), (6,)))
        (((1, 3), (2,)), ((5,), (6,)))
        (((2,), (1, 3)), ((5,), (6,)))
        (((2, 1), (3,)), ((5,), (6,)))
        (((2,), (3, 1)), ((5,), (6,)))
        (((2, 3), (1,)), ((5,), (6,)))
        (((3,), (1, 2)), ((5,), (6,)))
        (((3, 1), (2,)), ((5,), (6,)))
        (((3,), (2, 1)), ((5,), (6,)))
        (((3, 2), (1,)), ((5,), (6,)))
    """

    if ordered == "commutative":
        ordered = 11
    if ordered == "associative":
        ordered = None
    sm, bg = (A, B) if len(A) < len(B) else (B, A)
    for part in kbins(list(range(len(bg))), len(sm), ordered=ordered):
        if bg == B:
            yield tuple((a,) for a in A), partition(B, part)
        else:
            yield partition(A, part), tuple((b,) for b in B)

def partition(it, part):
    """ Partition a tuple/list into pieces defined by indices.

    Examples
    ========

    >>> from sympy.unify.core import partition
    >>> partition((10, 20, 30, 40), [[0, 1, 2], [3]])
    ((10, 20, 30), (40,))
    """
    return type(it)([index(it, ind) for ind in part])

def index(it, ind):
    """ Fancy indexing into an indexable iterable (tuple, list).

    Examples
    ========

    >>> from sympy.unify.core import index
    >>> index([10, 20, 30], (1, 2, 0))
    [20, 30, 10]
    """
    return type(it)([it[i] for i in ind])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\cli\base_command.py ===
"""Base Command class, and related routines"""

import logging
import logging.config
import optparse
import os
import sys
import traceback
from optparse import Values
from typing import List, Optional, Tuple

from pip._vendor.rich import reconfigure
from pip._vendor.rich import traceback as rich_traceback

from pip._internal.cli import cmdoptions
from pip._internal.cli.command_context import CommandContextMixIn
from pip._internal.cli.parser import ConfigOptionParser, UpdatingDefaultsHelpFormatter
from pip._internal.cli.status_codes import (
    ERROR,
    PREVIOUS_BUILD_DIR_ERROR,
    UNKNOWN_ERROR,
    VIRTUALENV_NOT_FOUND,
)
from pip._internal.exceptions import (
    BadCommand,
    CommandError,
    DiagnosticPipError,
    InstallationError,
    NetworkConnectionError,
    PreviousBuildDirError,
)
from pip._internal.utils.filesystem import check_path_owner
from pip._internal.utils.logging import BrokenStdoutLoggingError, setup_logging
from pip._internal.utils.misc import get_prog, normalize_path
from pip._internal.utils.temp_dir import TempDirectoryTypeRegistry as TempDirRegistry
from pip._internal.utils.temp_dir import global_tempdir_manager, tempdir_registry
from pip._internal.utils.virtualenv import running_under_virtualenv

__all__ = ["Command"]

logger = logging.getLogger(__name__)


class Command(CommandContextMixIn):
    usage: str = ""
    ignore_require_venv: bool = False

    def __init__(self, name: str, summary: str, isolated: bool = False) -> None:
        super().__init__()

        self.name = name
        self.summary = summary
        self.parser = ConfigOptionParser(
            usage=self.usage,
            prog=f"{get_prog()} {name}",
            formatter=UpdatingDefaultsHelpFormatter(),
            add_help_option=False,
            name=name,
            description=self.__doc__,
            isolated=isolated,
        )

        self.tempdir_registry: Optional[TempDirRegistry] = None

        # Commands should add options to this option group
        optgroup_name = f"{self.name.capitalize()} Options"
        self.cmd_opts = optparse.OptionGroup(self.parser, optgroup_name)

        # Add the general options
        gen_opts = cmdoptions.make_option_group(
            cmdoptions.general_group,
            self.parser,
        )
        self.parser.add_option_group(gen_opts)

        self.add_options()

    def add_options(self) -> None:
        pass

    def handle_pip_version_check(self, options: Values) -> None:
        """
        This is a no-op so that commands by default do not do the pip version
        check.
        """
        # Make sure we do the pip version check if the index_group options
        # are present.
        assert not hasattr(options, "no_index")

    def run(self, options: Values, args: List[str]) -> int:
        raise NotImplementedError

    def _run_wrapper(self, level_number: int, options: Values, args: List[str]) -> int:
        def _inner_run() -> int:
            try:
                return self.run(options, args)
            finally:
                self.handle_pip_version_check(options)

        if options.debug_mode:
            rich_traceback.install(show_locals=True)
            return _inner_run()

        try:
            status = _inner_run()
            assert isinstance(status, int)
            return status
        except DiagnosticPipError as exc:
            logger.error("%s", exc, extra={"rich": True})
            logger.debug("Exception information:", exc_info=True)

            return ERROR
        except PreviousBuildDirError as exc:
            logger.critical(str(exc))
            logger.debug("Exception information:", exc_info=True)

            return PREVIOUS_BUILD_DIR_ERROR
        except (
            InstallationError,
            BadCommand,
            NetworkConnectionError,
        ) as exc:
            logger.critical(str(exc))
            logger.debug("Exception information:", exc_info=True)

            return ERROR
        except CommandError as exc:
            logger.critical("%s", exc)
            logger.debug("Exception information:", exc_info=True)

            return ERROR
        except BrokenStdoutLoggingError:
            # Bypass our logger and write any remaining messages to
            # stderr because stdout no longer works.
            print("ERROR: Pipe to stdout was broken", file=sys.stderr)
            if level_number <= logging.DEBUG:
                traceback.print_exc(file=sys.stderr)

            return ERROR
        except KeyboardInterrupt:
            logger.critical("Operation cancelled by user")
            logger.debug("Exception information:", exc_info=True)

            return ERROR
        except BaseException:
            logger.critical("Exception:", exc_info=True)

            return UNKNOWN_ERROR

    def parse_args(self, args: List[str]) -> Tuple[Values, List[str]]:
        # factored out for testability
        return self.parser.parse_args(args)

    def main(self, args: List[str]) -> int:
        try:
            with self.main_context():
                return self._main(args)
        finally:
            logging.shutdown()

    def _main(self, args: List[str]) -> int:
        # We must initialize this before the tempdir manager, otherwise the
        # configuration would not be accessible by the time we clean up the
        # tempdir manager.
        self.tempdir_registry = self.enter_context(tempdir_registry())
        # Intentionally set as early as possible so globally-managed temporary
        # directories are available to the rest of the code.
        self.enter_context(global_tempdir_manager())

        options, args = self.parse_args(args)

        # Set verbosity so that it can be used elsewhere.
        self.verbosity = options.verbose - options.quiet
        if options.debug_mode:
            self.verbosity = 2

        reconfigure(no_color=options.no_color)
        level_number = setup_logging(
            verbosity=self.verbosity,
            no_color=options.no_color,
            user_log_file=options.log,
        )

        always_enabled_features = set(options.features_enabled) & set(
            cmdoptions.ALWAYS_ENABLED_FEATURES
        )
        if always_enabled_features:
            logger.warning(
                "The following features are always enabled: %s. ",
                ", ".join(sorted(always_enabled_features)),
            )

        # Make sure that the --python argument isn't specified after the
        # subcommand. We can tell, because if --python was specified,
        # we should only reach this point if we're running in the created
        # subprocess, which has the _PIP_RUNNING_IN_SUBPROCESS environment
        # variable set.
        if options.python and "_PIP_RUNNING_IN_SUBPROCESS" not in os.environ:
            logger.critical(
                "The --python option must be placed before the pip subcommand name"
            )
            sys.exit(ERROR)

        # TODO: Try to get these passing down from the command?
        #       without resorting to os.environ to hold these.
        #       This also affects isolated builds and it should.

        if options.no_input:
            os.environ["PIP_NO_INPUT"] = "1"

        if options.exists_action:
            os.environ["PIP_EXISTS_ACTION"] = " ".join(options.exists_action)

        if options.require_venv and not self.ignore_require_venv:
            # If a venv is required check if it can really be found
            if not running_under_virtualenv():
                logger.critical("Could not find an activated virtualenv (required).")
                sys.exit(VIRTUALENV_NOT_FOUND)

        if options.cache_dir:
            options.cache_dir = normalize_path(options.cache_dir)
            if not check_path_owner(options.cache_dir):
                logger.warning(
                    "The directory '%s' or its parent directory is not owned "
                    "or is not writable by the current user. The cache "
                    "has been disabled. Check the permissions and owner of "
                    "that directory. If executing pip with sudo, you should "
                    "use sudo's -H flag.",
                    options.cache_dir,
                )
                options.cache_dir = None

        return self._run_wrapper(level_number, options, args)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\plotting\plot_implicit.py ===
"""Implicit plotting module for SymPy.

Explanation
===========

The module implements a data series called ImplicitSeries which is used by
``Plot`` class to plot implicit plots for different backends. The module,
by default, implements plotting using interval arithmetic. It switches to a
fall back algorithm if the expression cannot be plotted using interval arithmetic.
It is also possible to specify to use the fall back algorithm for all plots.

Boolean combinations of expressions cannot be plotted by the fall back
algorithm.

See Also
========

sympy.plotting.plot

References
==========

.. [1] Jeffrey Allen Tupper. Reliable Two-Dimensional Graphing Methods for
Mathematical Formulae with Two Free Variables.

.. [2] Jeffrey Allen Tupper. Graphing Equations with Generalized Interval
Arithmetic. Master's thesis. University of Toronto, 1996

"""


from sympy.core.containers import Tuple
from sympy.core.symbol import (Dummy, Symbol)
from sympy.polys.polyutils import _sort_gens
from sympy.plotting.series import ImplicitSeries, _set_discretization_points
from sympy.plotting.plot import plot_factory
from sympy.utilities.decorator import doctest_depends_on
from sympy.utilities.iterables import flatten


__doctest_requires__ = {'plot_implicit': ['matplotlib']}


@doctest_depends_on(modules=('matplotlib',))
def plot_implicit(expr, x_var=None, y_var=None, adaptive=True, depth=0,
                  n=300, line_color="blue", show=True, **kwargs):
    """A plot function to plot implicit equations / inequalities.

    Arguments
    =========

    - expr : The equation / inequality that is to be plotted.
    - x_var (optional) : symbol to plot on x-axis or tuple giving symbol
      and range as ``(symbol, xmin, xmax)``
    - y_var (optional) : symbol to plot on y-axis or tuple giving symbol
      and range as ``(symbol, ymin, ymax)``

    If neither ``x_var`` nor ``y_var`` are given then the free symbols in the
    expression will be assigned in the order they are sorted.

    The following keyword arguments can also be used:

    - ``adaptive`` Boolean. The default value is set to True. It has to be
        set to False if you want to use a mesh grid.

    - ``depth`` integer. The depth of recursion for adaptive mesh grid.
        Default value is 0. Takes value in the range (0, 4).

    - ``n`` integer. The number of points if adaptive mesh grid is not
        used. Default value is 300. This keyword argument replaces ``points``,
        which should be considered deprecated.

    - ``show`` Boolean. Default value is True. If set to False, the plot will
        not be shown. See ``Plot`` for further information.

    - ``title`` string. The title for the plot.

    - ``xlabel`` string. The label for the x-axis

    - ``ylabel`` string. The label for the y-axis

    Aesthetics options:

    - ``line_color``: float or string. Specifies the color for the plot.
        See ``Plot`` to see how to set color for the plots.
        Default value is "Blue"

    plot_implicit, by default, uses interval arithmetic to plot functions. If
    the expression cannot be plotted using interval arithmetic, it defaults to
    a generating a contour using a mesh grid of fixed number of points. By
    setting adaptive to False, you can force plot_implicit to use the mesh
    grid. The mesh grid method can be effective when adaptive plotting using
    interval arithmetic, fails to plot with small line width.

    Examples
    ========

    Plot expressions:

    .. plot::
        :context: reset
        :format: doctest
        :include-source: True

        >>> from sympy import plot_implicit, symbols, Eq, And
        >>> x, y = symbols('x y')

    Without any ranges for the symbols in the expression:

    .. plot::
        :context: close-figs
        :format: doctest
        :include-source: True

        >>> p1 = plot_implicit(Eq(x**2 + y**2, 5))

    With the range for the symbols:

    .. plot::
        :context: close-figs
        :format: doctest
        :include-source: True

        >>> p2 = plot_implicit(
        ...     Eq(x**2 + y**2, 3), (x, -3, 3), (y, -3, 3))

    With depth of recursion as argument:

    .. plot::
        :context: close-figs
        :format: doctest
        :include-source: True

        >>> p3 = plot_implicit(
        ...     Eq(x**2 + y**2, 5), (x, -4, 4), (y, -4, 4), depth = 2)

    Using mesh grid and not using adaptive meshing:

    .. plot::
        :context: close-figs
        :format: doctest
        :include-source: True

        >>> p4 = plot_implicit(
        ...     Eq(x**2 + y**2, 5), (x, -5, 5), (y, -2, 2),
        ...     adaptive=False)

    Using mesh grid without using adaptive meshing with number of points
    specified:

    .. plot::
        :context: close-figs
        :format: doctest
        :include-source: True

        >>> p5 = plot_implicit(
        ...     Eq(x**2 + y**2, 5), (x, -5, 5), (y, -2, 2),
        ...     adaptive=False, n=400)

    Plotting regions:

    .. plot::
        :context: close-figs
        :format: doctest
        :include-source: True

        >>> p6 = plot_implicit(y > x**2)

    Plotting Using boolean conjunctions:

    .. plot::
        :context: close-figs
        :format: doctest
        :include-source: True

        >>> p7 = plot_implicit(And(y > x, y > -x))

    When plotting an expression with a single variable (y - 1, for example),
    specify the x or the y variable explicitly:

    .. plot::
        :context: close-figs
        :format: doctest
        :include-source: True

        >>> p8 = plot_implicit(y - 1, y_var=y)
        >>> p9 = plot_implicit(x - 1, x_var=x)
    """

    xyvar = [i for i in (x_var, y_var) if i is not None]
    free_symbols = expr.free_symbols
    range_symbols = Tuple(*flatten(xyvar)).free_symbols
    undeclared = free_symbols - range_symbols
    if len(free_symbols & range_symbols) > 2:
        raise NotImplementedError("Implicit plotting is not implemented for "
                                  "more than 2 variables")

    #Create default ranges if the range is not provided.
    default_range = Tuple(-5, 5)
    def _range_tuple(s):
        if isinstance(s, Symbol):
            return Tuple(s) + default_range
        if len(s) == 3:
            return Tuple(*s)
        raise ValueError('symbol or `(symbol, min, max)` expected but got %s' % s)

    if len(xyvar) == 0:
        xyvar = list(_sort_gens(free_symbols))
    var_start_end_x = _range_tuple(xyvar[0])
    x = var_start_end_x[0]
    if len(xyvar) != 2:
        if x in undeclared or not undeclared:
            xyvar.append(Dummy('f(%s)' % x.name))
        else:
            xyvar.append(undeclared.pop())
    var_start_end_y = _range_tuple(xyvar[1])

    kwargs = _set_discretization_points(kwargs, ImplicitSeries)
    series_argument = ImplicitSeries(
        expr, var_start_end_x, var_start_end_y,
        adaptive=adaptive, depth=depth,
        n=n, line_color=line_color)

    #set the x and y limits
    kwargs['xlim'] = tuple(float(x) for x in var_start_end_x[1:])
    kwargs['ylim'] = tuple(float(y) for y in var_start_end_y[1:])
    # set the x and y labels
    kwargs.setdefault('xlabel', var_start_end_x[0])
    kwargs.setdefault('ylabel', var_start_end_y[0])
    p = plot_factory(series_argument, **kwargs)
    if show:
        p.show()
    return p

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\_testing\_warnings.py ===
from __future__ import annotations

from contextlib import (
    contextmanager,
    nullcontext,
)
import inspect
import re
import sys
from typing import (
    TYPE_CHECKING,
    Literal,
    cast,
)
import warnings

from pandas.compat import PY311

if TYPE_CHECKING:
    from collections.abc import (
        Generator,
        Sequence,
    )


@contextmanager
def assert_produces_warning(
    expected_warning: type[Warning] | bool | tuple[type[Warning], ...] | None = Warning,
    filter_level: Literal[
        "error", "ignore", "always", "default", "module", "once"
    ] = "always",
    check_stacklevel: bool = True,
    raise_on_extra_warnings: bool = True,
    match: str | None = None,
) -> Generator[list[warnings.WarningMessage], None, None]:
    """
    Context manager for running code expected to either raise a specific warning,
    multiple specific warnings, or not raise any warnings. Verifies that the code
    raises the expected warning(s), and that it does not raise any other unexpected
    warnings. It is basically a wrapper around ``warnings.catch_warnings``.

    Parameters
    ----------
    expected_warning : {Warning, False, tuple[Warning, ...], None}, default Warning
        The type of Exception raised. ``exception.Warning`` is the base
        class for all warnings. To raise multiple types of exceptions,
        pass them as a tuple. To check that no warning is returned,
        specify ``False`` or ``None``.
    filter_level : str or None, default "always"
        Specifies whether warnings are ignored, displayed, or turned
        into errors.
        Valid values are:

        * "error" - turns matching warnings into exceptions
        * "ignore" - discard the warning
        * "always" - always emit a warning
        * "default" - print the warning the first time it is generated
          from each location
        * "module" - print the warning the first time it is generated
          from each module
        * "once" - print the warning the first time it is generated

    check_stacklevel : bool, default True
        If True, displays the line that called the function containing
        the warning to show were the function is called. Otherwise, the
        line that implements the function is displayed.
    raise_on_extra_warnings : bool, default True
        Whether extra warnings not of the type `expected_warning` should
        cause the test to fail.
    match : str, optional
        Match warning message.

    Examples
    --------
    >>> import warnings
    >>> with assert_produces_warning():
    ...     warnings.warn(UserWarning())
    ...
    >>> with assert_produces_warning(False):
    ...     warnings.warn(RuntimeWarning())
    ...
    Traceback (most recent call last):
        ...
    AssertionError: Caused unexpected warning(s): ['RuntimeWarning'].
    >>> with assert_produces_warning(UserWarning):
    ...     warnings.warn(RuntimeWarning())
    Traceback (most recent call last):
        ...
    AssertionError: Did not see expected warning of class 'UserWarning'.

    ..warn:: This is *not* thread-safe.
    """
    __tracebackhide__ = True

    with warnings.catch_warnings(record=True) as w:
        warnings.simplefilter(filter_level)
        try:
            yield w
        finally:
            if expected_warning:
                expected_warning = cast(type[Warning], expected_warning)
                _assert_caught_expected_warning(
                    caught_warnings=w,
                    expected_warning=expected_warning,
                    match=match,
                    check_stacklevel=check_stacklevel,
                )
            if raise_on_extra_warnings:
                _assert_caught_no_extra_warnings(
                    caught_warnings=w,
                    expected_warning=expected_warning,
                )


def maybe_produces_warning(warning: type[Warning], condition: bool, **kwargs):
    """
    Return a context manager that possibly checks a warning based on the condition
    """
    if condition:
        return assert_produces_warning(warning, **kwargs)
    else:
        return nullcontext()


def _assert_caught_expected_warning(
    *,
    caught_warnings: Sequence[warnings.WarningMessage],
    expected_warning: type[Warning],
    match: str | None,
    check_stacklevel: bool,
) -> None:
    """Assert that there was the expected warning among the caught warnings."""
    saw_warning = False
    matched_message = False
    unmatched_messages = []

    for actual_warning in caught_warnings:
        if issubclass(actual_warning.category, expected_warning):
            saw_warning = True

            if check_stacklevel:
                _assert_raised_with_correct_stacklevel(actual_warning)

            if match is not None:
                if re.search(match, str(actual_warning.message)):
                    matched_message = True
                else:
                    unmatched_messages.append(actual_warning.message)

    if not saw_warning:
        raise AssertionError(
            f"Did not see expected warning of class "
            f"{repr(expected_warning.__name__)}"
        )

    if match and not matched_message:
        raise AssertionError(
            f"Did not see warning {repr(expected_warning.__name__)} "
            f"matching '{match}'. The emitted warning messages are "
            f"{unmatched_messages}"
        )


def _assert_caught_no_extra_warnings(
    *,
    caught_warnings: Sequence[warnings.WarningMessage],
    expected_warning: type[Warning] | bool | tuple[type[Warning], ...] | None,
) -> None:
    """Assert that no extra warnings apart from the expected ones are caught."""
    extra_warnings = []

    for actual_warning in caught_warnings:
        if _is_unexpected_warning(actual_warning, expected_warning):
            # GH#38630 pytest.filterwarnings does not suppress these.
            if actual_warning.category == ResourceWarning:
                # GH 44732: Don't make the CI flaky by filtering SSL-related
                # ResourceWarning from dependencies
                if "unclosed <ssl.SSLSocket" in str(actual_warning.message):
                    continue
                # GH 44844: Matplotlib leaves font files open during the entire process
                # upon import. Don't make CI flaky if ResourceWarning raised
                # due to these open files.
                if any("matplotlib" in mod for mod in sys.modules):
                    continue
            if PY311 and actual_warning.category == EncodingWarning:
                # EncodingWarnings are checked in the CI
                # pyproject.toml errors on EncodingWarnings in pandas
                # Ignore EncodingWarnings from other libraries
                continue
            extra_warnings.append(
                (
                    actual_warning.category.__name__,
                    actual_warning.message,
                    actual_warning.filename,
                    actual_warning.lineno,
                )
            )

    if extra_warnings:
        raise AssertionError(f"Caused unexpected warning(s): {repr(extra_warnings)}")


def _is_unexpected_warning(
    actual_warning: warnings.WarningMessage,
    expected_warning: type[Warning] | bool | tuple[type[Warning], ...] | None,
) -> bool:
    """Check if the actual warning issued is unexpected."""
    if actual_warning and not expected_warning:
        return True
    expected_warning = cast(type[Warning], expected_warning)
    return bool(not issubclass(actual_warning.category, expected_warning))


def _assert_raised_with_correct_stacklevel(
    actual_warning: warnings.WarningMessage,
) -> None:
    # https://stackoverflow.com/questions/17407119/python-inspect-stack-is-slow
    frame = inspect.currentframe()
    for _ in range(4):
        frame = frame.f_back  # type: ignore[union-attr]
    try:
        caller_filename = inspect.getfile(frame)  # type: ignore[arg-type]
    finally:
        # See note in
        # https://docs.python.org/3/library/inspect.html#inspect.Traceback
        del frame
    msg = (
        "Warning not set with correct stacklevel. "
        f"File where warning is raised: {actual_warning.filename} != "
        f"{caller_filename}. Warning message: {actual_warning.message}"
    )
    assert actual_warning.filename == caller_filename, msg

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\highlighter.py ===
import re
from abc import ABC, abstractmethod
from typing import List, Union

from .text import Span, Text


def _combine_regex(*regexes: str) -> str:
    """Combine a number of regexes in to a single regex.

    Returns:
        str: New regex with all regexes ORed together.
    """
    return "|".join(regexes)


class Highlighter(ABC):
    """Abstract base class for highlighters."""

    def __call__(self, text: Union[str, Text]) -> Text:
        """Highlight a str or Text instance.

        Args:
            text (Union[str, ~Text]): Text to highlight.

        Raises:
            TypeError: If not called with text or str.

        Returns:
            Text: A test instance with highlighting applied.
        """
        if isinstance(text, str):
            highlight_text = Text(text)
        elif isinstance(text, Text):
            highlight_text = text.copy()
        else:
            raise TypeError(f"str or Text instance required, not {text!r}")
        self.highlight(highlight_text)
        return highlight_text

    @abstractmethod
    def highlight(self, text: Text) -> None:
        """Apply highlighting in place to text.

        Args:
            text (~Text): A text object highlight.
        """


class NullHighlighter(Highlighter):
    """A highlighter object that doesn't highlight.

    May be used to disable highlighting entirely.

    """

    def highlight(self, text: Text) -> None:
        """Nothing to do"""


class RegexHighlighter(Highlighter):
    """Applies highlighting from a list of regular expressions."""

    highlights: List[str] = []
    base_style: str = ""

    def highlight(self, text: Text) -> None:
        """Highlight :class:`rich.text.Text` using regular expressions.

        Args:
            text (~Text): Text to highlighted.

        """

        highlight_regex = text.highlight_regex
        for re_highlight in self.highlights:
            highlight_regex(re_highlight, style_prefix=self.base_style)


class ReprHighlighter(RegexHighlighter):
    """Highlights the text typically produced from ``__repr__`` methods."""

    base_style = "repr."
    highlights = [
        r"(?P<tag_start><)(?P<tag_name>[-\w.:|]*)(?P<tag_contents>[\w\W]*)(?P<tag_end>>)",
        r'(?P<attrib_name>[\w_]{1,50})=(?P<attrib_value>"?[\w_]+"?)?',
        r"(?P<brace>[][{}()])",
        _combine_regex(
            r"(?P<ipv4>[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3})",
            r"(?P<ipv6>([A-Fa-f0-9]{1,4}::?){1,7}[A-Fa-f0-9]{1,4})",
            r"(?P<eui64>(?:[0-9A-Fa-f]{1,2}-){7}[0-9A-Fa-f]{1,2}|(?:[0-9A-Fa-f]{1,2}:){7}[0-9A-Fa-f]{1,2}|(?:[0-9A-Fa-f]{4}\.){3}[0-9A-Fa-f]{4})",
            r"(?P<eui48>(?:[0-9A-Fa-f]{1,2}-){5}[0-9A-Fa-f]{1,2}|(?:[0-9A-Fa-f]{1,2}:){5}[0-9A-Fa-f]{1,2}|(?:[0-9A-Fa-f]{4}\.){2}[0-9A-Fa-f]{4})",
            r"(?P<uuid>[a-fA-F0-9]{8}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{12})",
            r"(?P<call>[\w.]*?)\(",
            r"\b(?P<bool_true>True)\b|\b(?P<bool_false>False)\b|\b(?P<none>None)\b",
            r"(?P<ellipsis>\.\.\.)",
            r"(?P<number_complex>(?<!\w)(?:\-?[0-9]+\.?[0-9]*(?:e[-+]?\d+?)?)(?:[-+](?:[0-9]+\.?[0-9]*(?:e[-+]?\d+)?))?j)",
            r"(?P<number>(?<!\w)\-?[0-9]+\.?[0-9]*(e[-+]?\d+?)?\b|0x[0-9a-fA-F]*)",
            r"(?P<path>\B(/[-\w._+]+)*\/)(?P<filename>[-\w._+]*)?",
            r"(?<![\\\w])(?P<str>b?'''.*?(?<!\\)'''|b?'.*?(?<!\\)'|b?\"\"\".*?(?<!\\)\"\"\"|b?\".*?(?<!\\)\")",
            r"(?P<url>(file|https|http|ws|wss)://[-0-9a-zA-Z$_+!`(),.?/;:&=%#~@]*)",
        ),
    ]


class JSONHighlighter(RegexHighlighter):
    """Highlights JSON"""

    # Captures the start and end of JSON strings, handling escaped quotes
    JSON_STR = r"(?<![\\\w])(?P<str>b?\".*?(?<!\\)\")"
    JSON_WHITESPACE = {" ", "\n", "\r", "\t"}

    base_style = "json."
    highlights = [
        _combine_regex(
            r"(?P<brace>[\{\[\(\)\]\}])",
            r"\b(?P<bool_true>true)\b|\b(?P<bool_false>false)\b|\b(?P<null>null)\b",
            r"(?P<number>(?<!\w)\-?[0-9]+\.?[0-9]*(e[\-\+]?\d+?)?\b|0x[0-9a-fA-F]*)",
            JSON_STR,
        ),
    ]

    def highlight(self, text: Text) -> None:
        super().highlight(text)

        # Additional work to handle highlighting JSON keys
        plain = text.plain
        append = text.spans.append
        whitespace = self.JSON_WHITESPACE
        for match in re.finditer(self.JSON_STR, plain):
            start, end = match.span()
            cursor = end
            while cursor < len(plain):
                char = plain[cursor]
                cursor += 1
                if char == ":":
                    append(Span(start, end, "json.key"))
                elif char in whitespace:
                    continue
                break


class ISO8601Highlighter(RegexHighlighter):
    """Highlights the ISO8601 date time strings.
    Regex reference: https://www.oreilly.com/library/view/regular-expressions-cookbook/9781449327453/ch04s07.html
    """

    base_style = "iso8601."
    highlights = [
        #
        # Dates
        #
        # Calendar month (e.g. 2008-08). The hyphen is required
        r"^(?P<year>[0-9]{4})-(?P<month>1[0-2]|0[1-9])$",
        # Calendar date w/o hyphens (e.g. 20080830)
        r"^(?P<date>(?P<year>[0-9]{4})(?P<month>1[0-2]|0[1-9])(?P<day>3[01]|0[1-9]|[12][0-9]))$",
        # Ordinal date (e.g. 2008-243). The hyphen is optional
        r"^(?P<date>(?P<year>[0-9]{4})-?(?P<day>36[0-6]|3[0-5][0-9]|[12][0-9]{2}|0[1-9][0-9]|00[1-9]))$",
        #
        # Weeks
        #
        # Week of the year (e.g., 2008-W35). The hyphen is optional
        r"^(?P<date>(?P<year>[0-9]{4})-?W(?P<week>5[0-3]|[1-4][0-9]|0[1-9]))$",
        # Week date (e.g., 2008-W35-6). The hyphens are optional
        r"^(?P<date>(?P<year>[0-9]{4})-?W(?P<week>5[0-3]|[1-4][0-9]|0[1-9])-?(?P<day>[1-7]))$",
        #
        # Times
        #
        # Hours and minutes (e.g., 17:21). The colon is optional
        r"^(?P<time>(?P<hour>2[0-3]|[01][0-9]):?(?P<minute>[0-5][0-9]))$",
        # Hours, minutes, and seconds w/o colons (e.g., 172159)
        r"^(?P<time>(?P<hour>2[0-3]|[01][0-9])(?P<minute>[0-5][0-9])(?P<second>[0-5][0-9]))$",
        # Time zone designator (e.g., Z, +07 or +07:00). The colons and the minutes are optional
        r"^(?P<timezone>(Z|[+-](?:2[0-3]|[01][0-9])(?::?(?:[0-5][0-9]))?))$",
        # Hours, minutes, and seconds with time zone designator (e.g., 17:21:59+07:00).
        # All the colons are optional. The minutes in the time zone designator are also optional
        r"^(?P<time>(?P<hour>2[0-3]|[01][0-9])(?P<minute>[0-5][0-9])(?P<second>[0-5][0-9]))(?P<timezone>Z|[+-](?:2[0-3]|[01][0-9])(?::?(?:[0-5][0-9]))?)$",
        #
        # Date and Time
        #
        # Calendar date with hours, minutes, and seconds (e.g., 2008-08-30 17:21:59 or 20080830 172159).
        # A space is required between the date and the time. The hyphens and colons are optional.
        # This regex matches dates and times that specify some hyphens or colons but omit others.
        # This does not follow ISO 8601
        r"^(?P<date>(?P<year>[0-9]{4})(?P<hyphen>-)?(?P<month>1[0-2]|0[1-9])(?(hyphen)-)(?P<day>3[01]|0[1-9]|[12][0-9])) (?P<time>(?P<hour>2[0-3]|[01][0-9])(?(hyphen):)(?P<minute>[0-5][0-9])(?(hyphen):)(?P<second>[0-5][0-9]))$",
        #
        # XML Schema dates and times
        #
        # Date, with optional time zone (e.g., 2008-08-30 or 2008-08-30+07:00).
        # Hyphens are required. This is the XML Schema 'date' type
        r"^(?P<date>(?P<year>-?(?:[1-9][0-9]*)?[0-9]{4})-(?P<month>1[0-2]|0[1-9])-(?P<day>3[01]|0[1-9]|[12][0-9]))(?P<timezone>Z|[+-](?:2[0-3]|[01][0-9]):[0-5][0-9])?$",
        # Time, with optional fractional seconds and time zone (e.g., 01:45:36 or 01:45:36.123+07:00).
        # There is no limit on the number of digits for the fractional seconds. This is the XML Schema 'time' type
        r"^(?P<time>(?P<hour>2[0-3]|[01][0-9]):(?P<minute>[0-5][0-9]):(?P<second>[0-5][0-9])(?P<frac>\.[0-9]+)?)(?P<timezone>Z|[+-](?:2[0-3]|[01][0-9]):[0-5][0-9])?$",
        # Date and time, with optional fractional seconds and time zone (e.g., 2008-08-30T01:45:36 or 2008-08-30T01:45:36.123Z).
        # This is the XML Schema 'dateTime' type
        r"^(?P<date>(?P<year>-?(?:[1-9][0-9]*)?[0-9]{4})-(?P<month>1[0-2]|0[1-9])-(?P<day>3[01]|0[1-9]|[12][0-9]))T(?P<time>(?P<hour>2[0-3]|[01][0-9]):(?P<minute>[0-5][0-9]):(?P<second>[0-5][0-9])(?P<ms>\.[0-9]+)?)(?P<timezone>Z|[+-](?:2[0-3]|[01][0-9]):[0-5][0-9])?$",
    ]


if __name__ == "__main__":  # pragma: no cover
    from .console import Console

    console = Console()
    console.print("[bold green]hello world![/bold green]")
    console.print("'[bold green]hello world![/bold green]'")

    console.print(" /foo")
    console.print("/foo/")
    console.print("/foo/bar")
    console.print("foo/bar/baz")

    console.print("/foo/bar/baz?foo=bar+egg&egg=baz")
    console.print("/foo/bar/baz/")
    console.print("/foo/bar/baz/egg")
    console.print("/foo/bar/baz/egg.py")
    console.print("/foo/bar/baz/egg.py word")
    console.print(" /foo/bar/baz/egg.py word")
    console.print("foo /foo/bar/baz/egg.py word")
    console.print("foo /foo/bar/ba._++z/egg+.py word")
    console.print("https://example.org?foo=bar#header")

    console.print(1234567.34)
    console.print(1 / 2)
    console.print(-1 / 123123123123)

    console.print(
        "127.0.1.1 bar 192.168.1.4 2001:0db8:85a3:0000:0000:8a2e:0370:7334 foo"
    )
    import json

    console.print_json(json.dumps(obj={"name": "apple", "count": 1}), indent=None)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\remote\errorhandler.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

from typing import Any, Dict, Type

from selenium.common.exceptions import (
    DetachedShadowRootException,
    ElementClickInterceptedException,
    ElementNotInteractableException,
    ElementNotSelectableException,
    ElementNotVisibleException,
    ImeActivationFailedException,
    ImeNotAvailableException,
    InsecureCertificateException,
    InvalidArgumentException,
    InvalidCookieDomainException,
    InvalidCoordinatesException,
    InvalidElementStateException,
    InvalidSelectorException,
    InvalidSessionIdException,
    JavascriptException,
    MoveTargetOutOfBoundsException,
    NoAlertPresentException,
    NoSuchCookieException,
    NoSuchElementException,
    NoSuchFrameException,
    NoSuchShadowRootException,
    NoSuchWindowException,
    ScreenshotException,
    SessionNotCreatedException,
    StaleElementReferenceException,
    TimeoutException,
    UnableToSetCookieException,
    UnexpectedAlertPresentException,
    UnknownMethodException,
    WebDriverException,
)


class ExceptionMapping:
    """
    :Maps each errorcode in ErrorCode object to corresponding exception
    Please refer to https://www.w3.org/TR/webdriver2/#errors for w3c specification
    """

    NO_SUCH_ELEMENT = NoSuchElementException
    NO_SUCH_FRAME = NoSuchFrameException
    NO_SUCH_SHADOW_ROOT = NoSuchShadowRootException
    STALE_ELEMENT_REFERENCE = StaleElementReferenceException
    ELEMENT_NOT_VISIBLE = ElementNotVisibleException
    INVALID_ELEMENT_STATE = InvalidElementStateException
    UNKNOWN_ERROR = WebDriverException
    ELEMENT_IS_NOT_SELECTABLE = ElementNotSelectableException
    JAVASCRIPT_ERROR = JavascriptException
    TIMEOUT = TimeoutException
    NO_SUCH_WINDOW = NoSuchWindowException
    INVALID_COOKIE_DOMAIN = InvalidCookieDomainException
    UNABLE_TO_SET_COOKIE = UnableToSetCookieException
    UNEXPECTED_ALERT_OPEN = UnexpectedAlertPresentException
    NO_ALERT_OPEN = NoAlertPresentException
    SCRIPT_TIMEOUT = TimeoutException
    IME_NOT_AVAILABLE = ImeNotAvailableException
    IME_ENGINE_ACTIVATION_FAILED = ImeActivationFailedException
    INVALID_SELECTOR = InvalidSelectorException
    SESSION_NOT_CREATED = SessionNotCreatedException
    MOVE_TARGET_OUT_OF_BOUNDS = MoveTargetOutOfBoundsException
    INVALID_XPATH_SELECTOR = InvalidSelectorException
    INVALID_XPATH_SELECTOR_RETURN_TYPER = InvalidSelectorException
    ELEMENT_NOT_INTERACTABLE = ElementNotInteractableException
    INSECURE_CERTIFICATE = InsecureCertificateException
    INVALID_ARGUMENT = InvalidArgumentException
    INVALID_COORDINATES = InvalidCoordinatesException
    INVALID_SESSION_ID = InvalidSessionIdException
    NO_SUCH_COOKIE = NoSuchCookieException
    UNABLE_TO_CAPTURE_SCREEN = ScreenshotException
    ELEMENT_CLICK_INTERCEPTED = ElementClickInterceptedException
    UNKNOWN_METHOD = UnknownMethodException
    DETACHED_SHADOW_ROOT = DetachedShadowRootException


class ErrorCode:
    """Error codes defined in the WebDriver wire protocol."""

    # Keep in sync with org.openqa.selenium.remote.ErrorCodes and errorcodes.h
    SUCCESS = 0
    NO_SUCH_ELEMENT = [7, "no such element"]
    NO_SUCH_FRAME = [8, "no such frame"]
    NO_SUCH_SHADOW_ROOT = ["no such shadow root"]
    UNKNOWN_COMMAND = [9, "unknown command"]
    STALE_ELEMENT_REFERENCE = [10, "stale element reference"]
    ELEMENT_NOT_VISIBLE = [11, "element not visible"]
    INVALID_ELEMENT_STATE = [12, "invalid element state"]
    UNKNOWN_ERROR = [13, "unknown error"]
    ELEMENT_IS_NOT_SELECTABLE = [15, "element not selectable"]
    JAVASCRIPT_ERROR = [17, "javascript error"]
    XPATH_LOOKUP_ERROR = [19, "invalid selector"]
    TIMEOUT = [21, "timeout"]
    NO_SUCH_WINDOW = [23, "no such window"]
    INVALID_COOKIE_DOMAIN = [24, "invalid cookie domain"]
    UNABLE_TO_SET_COOKIE = [25, "unable to set cookie"]
    UNEXPECTED_ALERT_OPEN = [26, "unexpected alert open"]
    NO_ALERT_OPEN = [27, "no such alert"]
    SCRIPT_TIMEOUT = [28, "script timeout"]
    INVALID_ELEMENT_COORDINATES = [29, "invalid element coordinates"]
    IME_NOT_AVAILABLE = [30, "ime not available"]
    IME_ENGINE_ACTIVATION_FAILED = [31, "ime engine activation failed"]
    INVALID_SELECTOR = [32, "invalid selector"]
    SESSION_NOT_CREATED = [33, "session not created"]
    MOVE_TARGET_OUT_OF_BOUNDS = [34, "move target out of bounds"]
    INVALID_XPATH_SELECTOR = [51, "invalid selector"]
    INVALID_XPATH_SELECTOR_RETURN_TYPER = [52, "invalid selector"]

    ELEMENT_NOT_INTERACTABLE = [60, "element not interactable"]
    INSECURE_CERTIFICATE = ["insecure certificate"]
    INVALID_ARGUMENT = [61, "invalid argument"]
    INVALID_COORDINATES = ["invalid coordinates"]
    INVALID_SESSION_ID = ["invalid session id"]
    NO_SUCH_COOKIE = [62, "no such cookie"]
    UNABLE_TO_CAPTURE_SCREEN = [63, "unable to capture screen"]
    ELEMENT_CLICK_INTERCEPTED = [64, "element click intercepted"]
    UNKNOWN_METHOD = ["unknown method exception"]
    DETACHED_SHADOW_ROOT = [65, "detached shadow root"]

    METHOD_NOT_ALLOWED = [405, "unsupported operation"]


class ErrorHandler:
    """Handles errors returned by the WebDriver server."""

    def check_response(self, response: Dict[str, Any]) -> None:
        """Checks that a JSON response from the WebDriver does not have an
        error.

        :Args:
         - response - The JSON response from the WebDriver server as a dictionary
           object.

        :Raises: If the response contains an error message.
        """
        status = response.get("status", None)
        if not status or status == ErrorCode.SUCCESS:
            return
        value = None
        message = response.get("message", "")
        screen: str = response.get("screen", "")
        stacktrace = None
        if isinstance(status, int):
            value_json = response.get("value", None)
            if value_json and isinstance(value_json, str):
                import json

                try:
                    value = json.loads(value_json)
                    if len(value) == 1:
                        value = value["value"]
                    status = value.get("error", None)
                    if not status:
                        status = value.get("status", ErrorCode.UNKNOWN_ERROR)
                        message = value.get("value") or value.get("message")
                        if not isinstance(message, str):
                            value = message
                            message = message.get("message")
                    else:
                        message = value.get("message", None)
                except ValueError:
                    pass

        exception_class: Type[WebDriverException]
        e = ErrorCode()
        error_codes = [item for item in dir(e) if not item.startswith("__")]
        for error_code in error_codes:
            error_info = getattr(ErrorCode, error_code)
            if isinstance(error_info, list) and status in error_info:
                exception_class = getattr(ExceptionMapping, error_code, WebDriverException)
                break
        else:
            exception_class = WebDriverException

        if not value:
            value = response["value"]
        if isinstance(value, str):
            raise exception_class(value)
        if message == "" and "message" in value:
            message = value["message"]

        screen = None  # type: ignore[assignment]
        if "screen" in value:
            screen = value["screen"]

        stacktrace = None
        st_value = value.get("stackTrace") or value.get("stacktrace")
        if st_value:
            if isinstance(st_value, str):
                stacktrace = st_value.split("\n")
            else:
                stacktrace = []
                try:
                    for frame in st_value:
                        line = frame.get("lineNumber", "")
                        file = frame.get("fileName", "<anonymous>")
                        if line:
                            file = f"{file}:{line}"
                        meth = frame.get("methodName", "<anonymous>")
                        if "className" in frame:
                            meth = f"{frame['className']}.{meth}"
                        msg = "    at %s (%s)"
                        msg = msg % (meth, file)
                        stacktrace.append(msg)
                except TypeError:
                    pass
        if exception_class == UnexpectedAlertPresentException:
            alert_text = None
            if "data" in value:
                alert_text = value["data"].get("text")
            elif "alert" in value:
                alert_text = value["alert"].get("text")
            raise exception_class(message, screen, stacktrace, alert_text)  # type: ignore[call-arg]  # mypy is not smart enough here
        raise exception_class(message, screen, stacktrace)