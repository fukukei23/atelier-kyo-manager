
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\tools\pytorch_export_helpers.py ===
# Copyright (c) Microsoft Corporation. All rights reserved.
# Licensed under the MIT License.

import inspect
from collections import abc

import torch


def _parse_inputs_for_onnx_export(all_input_parameters, inputs, kwargs):
    # extracted from https://github.com/microsoft/onnxruntime/blob/239c6ad3f021ff7cc2e6247eb074bd4208dc11e2/orttraining/orttraining/python/training/ortmodule/_io.py#L433

    def _add_input(name, input):
        """Returns number of expanded inputs that _add_input processed"""

        if input is None:
            # Drop all None inputs and return 0.
            return 0

        num_expanded_non_none_inputs = 0
        if isinstance(input, abc.Sequence):
            # If the input is a sequence (like a list), expand the list so that
            # each element of the list is an input by itself.
            for i, val in enumerate(input):
                # Name each input with the index appended to the original name of the
                # argument.
                num_expanded_non_none_inputs += _add_input(f"{name}_{i}", val)

            # Return here since the list by itself is not a valid input.
            # All the elements of the list have already been added as inputs individually.
            return num_expanded_non_none_inputs
        elif isinstance(input, abc.Mapping):
            # If the input is a mapping (like a dict), expand the dict so that
            # each element of the dict is an input by itself.
            for key, val in input.items():
                num_expanded_non_none_inputs += _add_input(f"{name}_{key}", val)

            # Return here since the dict by itself is not a valid input.
            # All the elements of the dict have already been added as inputs individually.
            return num_expanded_non_none_inputs

        # InputInfo should contain all the names irrespective of whether they are
        # a part of the onnx graph or not.
        input_names.append(name)

        # A single input non none input was processed, return 1
        return 1

    input_names = []
    var_positional_idx = 0
    num_expanded_non_none_positional_inputs = 0

    for input_idx, input_parameter in enumerate(all_input_parameters):
        if input_parameter.kind == inspect.Parameter.VAR_POSITIONAL:
            # VAR_POSITIONAL parameter carries all *args parameters from original forward method
            for args_i in range(input_idx, len(inputs)):
                name = f"{input_parameter.name}_{var_positional_idx}"
                var_positional_idx += 1
                inp = inputs[args_i]
                num_expanded_non_none_positional_inputs += _add_input(name, inp)
        elif (
            input_parameter.kind == inspect.Parameter.POSITIONAL_ONLY
            or input_parameter.kind == inspect.Parameter.POSITIONAL_OR_KEYWORD
            or input_parameter.kind == inspect.Parameter.KEYWORD_ONLY
        ):
            # All positional non-*args and non-**kwargs are processed here
            name = input_parameter.name
            inp = None
            input_idx += var_positional_idx  # noqa: PLW2901
            is_positional = True
            if input_idx < len(inputs) and inputs[input_idx] is not None:
                inp = inputs[input_idx]
            elif name in kwargs and kwargs[name] is not None:
                inp = kwargs[name]
                is_positional = False
            num_expanded_non_none_inputs_local = _add_input(name, inp)
            if is_positional:
                num_expanded_non_none_positional_inputs += num_expanded_non_none_inputs_local
        elif input_parameter.kind == inspect.Parameter.VAR_KEYWORD:
            # **kwargs is always the last argument of forward()
            for name, inp in kwargs.items():
                if name not in input_names:
                    _add_input(name, inp)

    return input_names


def _flatten_module_input(names, args, kwargs):
    """Flatten args and kwargs in a single tuple of tensors."""
    # extracted from https://github.com/microsoft/onnxruntime/blob/239c6ad3f021ff7cc2e6247eb074bd4208dc11e2/orttraining/orttraining/python/training/ortmodule/_io.py#L110

    def is_primitive_type(value):
        return type(value) in {int, bool, float}

    def to_tensor(value):
        return torch.tensor(value)

    ret = [to_tensor(arg) if is_primitive_type(arg) else arg for arg in args]
    ret += [
        to_tensor(kwargs[name]) if is_primitive_type(kwargs[name]) else kwargs[name] for name in names if name in kwargs
    ]

    # if kwargs is empty, append an empty dictionary at the end of the sample inputs to make exporter
    # happy. This is because the exporter is confused with kwargs and dictionary inputs otherwise.
    if not kwargs:
        ret.append({})

    return tuple(ret)


def infer_input_info(module: torch.nn.Module, *inputs, **kwargs):
    """
    Infer the input names and order from the arguments used to execute a PyTorch module for usage exporting
    the model via torch.onnx.export.
    Assumes model is on CPU. Use `module.to(torch.device('cpu'))` if it isn't.

    Example usage:
    input_names, inputs_as_tuple = infer_input_info(module, ...)
    torch.onnx.export(module, inputs_as_type, 'model.onnx', input_names=input_names, output_names=[...], ...)

    :param module: Module
    :param inputs: Positional inputs
    :param kwargs: Keyword argument inputs
    :return: Tuple of ordered input names and input values. These can be used directly with torch.onnx.export as the
            `input_names` and `inputs` arguments.
    """
    module_parameters = inspect.signature(module.forward).parameters.values()
    input_names = _parse_inputs_for_onnx_export(module_parameters, inputs, kwargs)
    inputs_as_tuple = _flatten_module_input(input_names, inputs, kwargs)

    return input_names, inputs_as_tuple

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\status.py ===
from types import TracebackType
from typing import Optional, Type

from .console import Console, RenderableType
from .jupyter import JupyterMixin
from .live import Live
from .spinner import Spinner
from .style import StyleType


class Status(JupyterMixin):
    """Displays a status indicator with a 'spinner' animation.

    Args:
        status (RenderableType): A status renderable (str or Text typically).
        console (Console, optional): Console instance to use, or None for global console. Defaults to None.
        spinner (str, optional): Name of spinner animation (see python -m rich.spinner). Defaults to "dots".
        spinner_style (StyleType, optional): Style of spinner. Defaults to "status.spinner".
        speed (float, optional): Speed factor for spinner animation. Defaults to 1.0.
        refresh_per_second (float, optional): Number of refreshes per second. Defaults to 12.5.
    """

    def __init__(
        self,
        status: RenderableType,
        *,
        console: Optional[Console] = None,
        spinner: str = "dots",
        spinner_style: StyleType = "status.spinner",
        speed: float = 1.0,
        refresh_per_second: float = 12.5,
    ):
        self.status = status
        self.spinner_style = spinner_style
        self.speed = speed
        self._spinner = Spinner(spinner, text=status, style=spinner_style, speed=speed)
        self._live = Live(
            self.renderable,
            console=console,
            refresh_per_second=refresh_per_second,
            transient=True,
        )

    @property
    def renderable(self) -> Spinner:
        return self._spinner

    @property
    def console(self) -> "Console":
        """Get the Console used by the Status objects."""
        return self._live.console

    def update(
        self,
        status: Optional[RenderableType] = None,
        *,
        spinner: Optional[str] = None,
        spinner_style: Optional[StyleType] = None,
        speed: Optional[float] = None,
    ) -> None:
        """Update status.

        Args:
            status (Optional[RenderableType], optional): New status renderable or None for no change. Defaults to None.
            spinner (Optional[str], optional): New spinner or None for no change. Defaults to None.
            spinner_style (Optional[StyleType], optional): New spinner style or None for no change. Defaults to None.
            speed (Optional[float], optional): Speed factor for spinner animation or None for no change. Defaults to None.
        """
        if status is not None:
            self.status = status
        if spinner_style is not None:
            self.spinner_style = spinner_style
        if speed is not None:
            self.speed = speed
        if spinner is not None:
            self._spinner = Spinner(
                spinner, text=self.status, style=self.spinner_style, speed=self.speed
            )
            self._live.update(self.renderable, refresh=True)
        else:
            self._spinner.update(
                text=self.status, style=self.spinner_style, speed=self.speed
            )

    def start(self) -> None:
        """Start the status animation."""
        self._live.start()

    def stop(self) -> None:
        """Stop the spinner animation."""
        self._live.stop()

    def __rich__(self) -> RenderableType:
        return self.renderable

    def __enter__(self) -> "Status":
        self.start()
        return self

    def __exit__(
        self,
        exc_type: Optional[Type[BaseException]],
        exc_val: Optional[BaseException],
        exc_tb: Optional[TracebackType],
    ) -> None:
        self.stop()


if __name__ == "__main__":  # pragma: no cover
    from time import sleep

    from .console import Console

    console = Console()
    with console.status("[magenta]Covid detector booting up") as status:
        sleep(3)
        console.log("Importing advanced AI")
        sleep(3)
        console.log("Advanced Covid AI Ready")
        sleep(3)
        status.update(status="[bold blue] Scanning for Covid", spinner="earth")
        sleep(3)
        console.log("Found 10,000,000,000 copies of Covid32.exe")
        sleep(3)
        status.update(
            status="[bold red]Moving Covid32.exe to Trash",
            spinner="bouncingBall",
            spinner_style="yellow",
        )
        sleep(5)
    console.print("[bold green]Covid deleted successfully")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\core\multidimensional.py ===
"""
Provides functionality for multidimensional usage of scalar-functions.

Read the vectorize docstring for more details.
"""

from functools import wraps


def apply_on_element(f, args, kwargs, n):
    """
    Returns a structure with the same dimension as the specified argument,
    where each basic element is replaced by the function f applied on it. All
    other arguments stay the same.
    """
    # Get the specified argument.
    if isinstance(n, int):
        structure = args[n]
        is_arg = True
    elif isinstance(n, str):
        structure = kwargs[n]
        is_arg = False

    # Define reduced function that is only dependent on the specified argument.
    def f_reduced(x):
        if hasattr(x, "__iter__"):
            return list(map(f_reduced, x))
        else:
            if is_arg:
                args[n] = x
            else:
                kwargs[n] = x
            return f(*args, **kwargs)

    # f_reduced will call itself recursively so that in the end f is applied to
    # all basic elements.
    return list(map(f_reduced, structure))


def iter_copy(structure):
    """
    Returns a copy of an iterable object (also copying all embedded iterables).
    """
    return [iter_copy(i) if hasattr(i, "__iter__") else i for i in structure]


def structure_copy(structure):
    """
    Returns a copy of the given structure (numpy-array, list, iterable, ..).
    """
    if hasattr(structure, "copy"):
        return structure.copy()
    return iter_copy(structure)


class vectorize:
    """
    Generalizes a function taking scalars to accept multidimensional arguments.

    Examples
    ========

    >>> from sympy import vectorize, diff, sin, symbols, Function
    >>> x, y, z = symbols('x y z')
    >>> f, g, h = list(map(Function, 'fgh'))

    >>> @vectorize(0)
    ... def vsin(x):
    ...     return sin(x)

    >>> vsin([1, x, y])
    [sin(1), sin(x), sin(y)]

    >>> @vectorize(0, 1)
    ... def vdiff(f, y):
    ...     return diff(f, y)

    >>> vdiff([f(x, y, z), g(x, y, z), h(x, y, z)], [x, y, z])
    [[Derivative(f(x, y, z), x), Derivative(f(x, y, z), y), Derivative(f(x, y, z), z)], [Derivative(g(x, y, z), x), Derivative(g(x, y, z), y), Derivative(g(x, y, z), z)], [Derivative(h(x, y, z), x), Derivative(h(x, y, z), y), Derivative(h(x, y, z), z)]]
    """
    def __init__(self, *mdargs):
        """
        The given numbers and strings characterize the arguments that will be
        treated as data structures, where the decorated function will be applied
        to every single element.
        If no argument is given, everything is treated multidimensional.
        """
        for a in mdargs:
            if not isinstance(a, (int, str)):
                raise TypeError("a is of invalid type")
        self.mdargs = mdargs

    def __call__(self, f):
        """
        Returns a wrapper for the one-dimensional function that can handle
        multidimensional arguments.
        """
        @wraps(f)
        def wrapper(*args, **kwargs):
            # Get arguments that should be treated multidimensional
            if self.mdargs:
                mdargs = self.mdargs
            else:
                mdargs = range(len(args)) + kwargs.keys()

            arglength = len(args)

            for n in mdargs:
                if isinstance(n, int):
                    if n >= arglength:
                        continue
                    entry = args[n]
                    is_arg = True
                elif isinstance(n, str):
                    try:
                        entry = kwargs[n]
                    except KeyError:
                        continue
                    is_arg = False
                if hasattr(entry, "__iter__"):
                    # Create now a copy of the given array and manipulate then
                    # the entries directly.
                    if is_arg:
                        args = list(args)
                        args[n] = structure_copy(entry)
                    else:
                        kwargs[n] = structure_copy(entry)
                    result = apply_on_element(wrapper, args, kwargs, n)
                    return result
            return f(*args, **kwargs)
        return wrapper

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\jinja2\nativetypes.py ===
import typing as t
from ast import literal_eval
from ast import parse
from itertools import chain
from itertools import islice
from types import GeneratorType

from . import nodes
from .compiler import CodeGenerator
from .compiler import Frame
from .compiler import has_safe_repr
from .environment import Environment
from .environment import Template


def native_concat(values: t.Iterable[t.Any]) -> t.Optional[t.Any]:
    """Return a native Python type from the list of compiled nodes. If
    the result is a single node, its value is returned. Otherwise, the
    nodes are concatenated as strings. If the result can be parsed with
    :func:`ast.literal_eval`, the parsed value is returned. Otherwise,
    the string is returned.

    :param values: Iterable of outputs to concatenate.
    """
    head = list(islice(values, 2))

    if not head:
        return None

    if len(head) == 1:
        raw = head[0]
        if not isinstance(raw, str):
            return raw
    else:
        if isinstance(values, GeneratorType):
            values = chain(head, values)
        raw = "".join([str(v) for v in values])

    try:
        return literal_eval(
            # In Python 3.10+ ast.literal_eval removes leading spaces/tabs
            # from the given string. For backwards compatibility we need to
            # parse the string ourselves without removing leading spaces/tabs.
            parse(raw, mode="eval")
        )
    except (ValueError, SyntaxError, MemoryError):
        return raw


class NativeCodeGenerator(CodeGenerator):
    """A code generator which renders Python types by not adding
    ``str()`` around output nodes.
    """

    @staticmethod
    def _default_finalize(value: t.Any) -> t.Any:
        return value

    def _output_const_repr(self, group: t.Iterable[t.Any]) -> str:
        return repr("".join([str(v) for v in group]))

    def _output_child_to_const(
        self, node: nodes.Expr, frame: Frame, finalize: CodeGenerator._FinalizeInfo
    ) -> t.Any:
        const = node.as_const(frame.eval_ctx)

        if not has_safe_repr(const):
            raise nodes.Impossible()

        if isinstance(node, nodes.TemplateData):
            return const

        return finalize.const(const)  # type: ignore

    def _output_child_pre(
        self, node: nodes.Expr, frame: Frame, finalize: CodeGenerator._FinalizeInfo
    ) -> None:
        if finalize.src is not None:
            self.write(finalize.src)

    def _output_child_post(
        self, node: nodes.Expr, frame: Frame, finalize: CodeGenerator._FinalizeInfo
    ) -> None:
        if finalize.src is not None:
            self.write(")")


class NativeEnvironment(Environment):
    """An environment that renders templates to native Python types."""

    code_generator_class = NativeCodeGenerator
    concat = staticmethod(native_concat)  # type: ignore


class NativeTemplate(Template):
    environment_class = NativeEnvironment

    def render(self, *args: t.Any, **kwargs: t.Any) -> t.Any:
        """Render the template to produce a native Python type. If the
        result is a single node, its value is returned. Otherwise, the
        nodes are concatenated as strings. If the result can be parsed
        with :func:`ast.literal_eval`, the parsed value is returned.
        Otherwise, the string is returned.
        """
        ctx = self.new_context(dict(*args, **kwargs))

        try:
            return self.environment_class.concat(  # type: ignore
                self.root_render_func(ctx)
            )
        except Exception:
            return self.environment.handle_exception()

    async def render_async(self, *args: t.Any, **kwargs: t.Any) -> t.Any:
        if not self.environment.is_async:
            raise RuntimeError(
                "The environment was not created with async mode enabled."
            )

        ctx = self.new_context(dict(*args, **kwargs))

        try:
            return self.environment_class.concat(  # type: ignore
                [n async for n in self.root_render_func(ctx)]  # type: ignore
            )
        except Exception:
            return self.environment.handle_exception()


NativeEnvironment.template_class = NativeTemplate

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\io\feather_format.py ===
""" feather-format compat """
from __future__ import annotations

from typing import (
    TYPE_CHECKING,
    Any,
)

from pandas._config import using_string_dtype

from pandas._libs import lib
from pandas.compat._optional import import_optional_dependency
from pandas.util._decorators import doc
from pandas.util._validators import check_dtype_backend

from pandas.core.api import DataFrame
from pandas.core.shared_docs import _shared_docs

from pandas.io._util import arrow_table_to_pandas
from pandas.io.common import get_handle

if TYPE_CHECKING:
    from collections.abc import (
        Hashable,
        Sequence,
    )

    from pandas._typing import (
        DtypeBackend,
        FilePath,
        ReadBuffer,
        StorageOptions,
        WriteBuffer,
    )


@doc(storage_options=_shared_docs["storage_options"])
def to_feather(
    df: DataFrame,
    path: FilePath | WriteBuffer[bytes],
    storage_options: StorageOptions | None = None,
    **kwargs: Any,
) -> None:
    """
    Write a DataFrame to the binary Feather format.

    Parameters
    ----------
    df : DataFrame
    path : str, path object, or file-like object
    {storage_options}
    **kwargs :
        Additional keywords passed to `pyarrow.feather.write_feather`.

    """
    import_optional_dependency("pyarrow")
    from pyarrow import feather

    if not isinstance(df, DataFrame):
        raise ValueError("feather only support IO with DataFrames")

    with get_handle(
        path, "wb", storage_options=storage_options, is_text=False
    ) as handles:
        feather.write_feather(df, handles.handle, **kwargs)


@doc(storage_options=_shared_docs["storage_options"])
def read_feather(
    path: FilePath | ReadBuffer[bytes],
    columns: Sequence[Hashable] | None = None,
    use_threads: bool = True,
    storage_options: StorageOptions | None = None,
    dtype_backend: DtypeBackend | lib.NoDefault = lib.no_default,
) -> DataFrame:
    """
    Load a feather-format object from the file path.

    Parameters
    ----------
    path : str, path object, or file-like object
        String, path object (implementing ``os.PathLike[str]``), or file-like
        object implementing a binary ``read()`` function. The string could be a URL.
        Valid URL schemes include http, ftp, s3, and file. For file URLs, a host is
        expected. A local file could be: ``file://localhost/path/to/table.feather``.
    columns : sequence, default None
        If not provided, all columns are read.
    use_threads : bool, default True
        Whether to parallelize reading using multiple threads.
    {storage_options}

    dtype_backend : {{'numpy_nullable', 'pyarrow'}}, default 'numpy_nullable'
        Back-end data type applied to the resultant :class:`DataFrame`
        (still experimental). Behaviour is as follows:

        * ``"numpy_nullable"``: returns nullable-dtype-backed :class:`DataFrame`
          (default).
        * ``"pyarrow"``: returns pyarrow-backed nullable :class:`ArrowDtype`
          DataFrame.

        .. versionadded:: 2.0

    Returns
    -------
    type of object stored in file

    Examples
    --------
    >>> df = pd.read_feather("path/to/file.feather")  # doctest: +SKIP
    """
    import_optional_dependency("pyarrow")
    from pyarrow import feather

    # import utils to register the pyarrow extension types
    import pandas.core.arrays.arrow.extension_types  # pyright: ignore[reportUnusedImport] # noqa: F401

    check_dtype_backend(dtype_backend)

    with get_handle(
        path, "rb", storage_options=storage_options, is_text=False
    ) as handles:
        if dtype_backend is lib.no_default and not using_string_dtype():
            return feather.read_feather(
                handles.handle, columns=columns, use_threads=bool(use_threads)
            )

        pa_table = feather.read_table(
            handles.handle, columns=columns, use_threads=bool(use_threads)
        )
        return arrow_table_to_pandas(pa_table, dtype_backend=dtype_backend)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\rule.py ===
from typing import Union

from .align import AlignMethod
from .cells import cell_len, set_cell_size
from .console import Console, ConsoleOptions, RenderResult
from .jupyter import JupyterMixin
from .measure import Measurement
from .style import Style
from .text import Text


class Rule(JupyterMixin):
    """A console renderable to draw a horizontal rule (line).

    Args:
        title (Union[str, Text], optional): Text to render in the rule. Defaults to "".
        characters (str, optional): Character(s) used to draw the line. Defaults to "─".
        style (StyleType, optional): Style of Rule. Defaults to "rule.line".
        end (str, optional): Character at end of Rule. defaults to "\\\\n"
        align (str, optional): How to align the title, one of "left", "center", or "right". Defaults to "center".
    """

    def __init__(
        self,
        title: Union[str, Text] = "",
        *,
        characters: str = "─",
        style: Union[str, Style] = "rule.line",
        end: str = "\n",
        align: AlignMethod = "center",
    ) -> None:
        if cell_len(characters) < 1:
            raise ValueError(
                "'characters' argument must have a cell width of at least 1"
            )
        if align not in ("left", "center", "right"):
            raise ValueError(
                f'invalid value for align, expected "left", "center", "right" (not {align!r})'
            )
        self.title = title
        self.characters = characters
        self.style = style
        self.end = end
        self.align = align

    def __repr__(self) -> str:
        return f"Rule({self.title!r}, {self.characters!r})"

    def __rich_console__(
        self, console: Console, options: ConsoleOptions
    ) -> RenderResult:
        width = options.max_width

        characters = (
            "-"
            if (options.ascii_only and not self.characters.isascii())
            else self.characters
        )

        chars_len = cell_len(characters)
        if not self.title:
            yield self._rule_line(chars_len, width)
            return

        if isinstance(self.title, Text):
            title_text = self.title
        else:
            title_text = console.render_str(self.title, style="rule.text")

        title_text.plain = title_text.plain.replace("\n", " ")
        title_text.expand_tabs()

        required_space = 4 if self.align == "center" else 2
        truncate_width = max(0, width - required_space)
        if not truncate_width:
            yield self._rule_line(chars_len, width)
            return

        rule_text = Text(end=self.end)
        if self.align == "center":
            title_text.truncate(truncate_width, overflow="ellipsis")
            side_width = (width - cell_len(title_text.plain)) // 2
            left = Text(characters * (side_width // chars_len + 1))
            left.truncate(side_width - 1)
            right_length = width - cell_len(left.plain) - cell_len(title_text.plain)
            right = Text(characters * (side_width // chars_len + 1))
            right.truncate(right_length)
            rule_text.append(left.plain + " ", self.style)
            rule_text.append(title_text)
            rule_text.append(" " + right.plain, self.style)
        elif self.align == "left":
            title_text.truncate(truncate_width, overflow="ellipsis")
            rule_text.append(title_text)
            rule_text.append(" ")
            rule_text.append(characters * (width - rule_text.cell_len), self.style)
        elif self.align == "right":
            title_text.truncate(truncate_width, overflow="ellipsis")
            rule_text.append(characters * (width - title_text.cell_len - 1), self.style)
            rule_text.append(" ")
            rule_text.append(title_text)

        rule_text.plain = set_cell_size(rule_text.plain, width)
        yield rule_text

    def _rule_line(self, chars_len: int, width: int) -> Text:
        rule_text = Text(self.characters * ((width // chars_len) + 1), self.style)
        rule_text.truncate(width)
        rule_text.plain = set_cell_size(rule_text.plain, width)
        return rule_text

    def __rich_measure__(
        self, console: Console, options: ConsoleOptions
    ) -> Measurement:
        return Measurement(1, 1)


if __name__ == "__main__":  # pragma: no cover
    import sys

    from pip._vendor.rich.console import Console

    try:
        text = sys.argv[1]
    except IndexError:
        text = "Hello, World"
    console = Console()
    console.print(Rule(title=text))

    console = Console()
    console.print(Rule("foo"), width=4)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\urllib3\contrib\ntlmpool.py ===
"""
NTLM authenticating pool, contributed by erikcederstran

Issue #10, see: http://code.google.com/p/urllib3/issues/detail?id=10
"""
from __future__ import absolute_import

import warnings
from logging import getLogger

from ntlm import ntlm

from .. import HTTPSConnectionPool
from ..packages.six.moves.http_client import HTTPSConnection

warnings.warn(
    "The 'urllib3.contrib.ntlmpool' module is deprecated and will be removed "
    "in urllib3 v2.0 release, urllib3 is not able to support it properly due "
    "to reasons listed in issue: https://github.com/urllib3/urllib3/issues/2282. "
    "If you are a user of this module please comment in the mentioned issue.",
    DeprecationWarning,
)

log = getLogger(__name__)


class NTLMConnectionPool(HTTPSConnectionPool):
    """
    Implements an NTLM authentication version of an urllib3 connection pool
    """

    scheme = "https"

    def __init__(self, user, pw, authurl, *args, **kwargs):
        """
        authurl is a random URL on the server that is protected by NTLM.
        user is the Windows user, probably in the DOMAIN\\username format.
        pw is the password for the user.
        """
        super(NTLMConnectionPool, self).__init__(*args, **kwargs)
        self.authurl = authurl
        self.rawuser = user
        user_parts = user.split("\\", 1)
        self.domain = user_parts[0].upper()
        self.user = user_parts[1]
        self.pw = pw

    def _new_conn(self):
        # Performs the NTLM handshake that secures the connection. The socket
        # must be kept open while requests are performed.
        self.num_connections += 1
        log.debug(
            "Starting NTLM HTTPS connection no. %d: https://%s%s",
            self.num_connections,
            self.host,
            self.authurl,
        )

        headers = {"Connection": "Keep-Alive"}
        req_header = "Authorization"
        resp_header = "www-authenticate"

        conn = HTTPSConnection(host=self.host, port=self.port)

        # Send negotiation message
        headers[req_header] = "NTLM %s" % ntlm.create_NTLM_NEGOTIATE_MESSAGE(
            self.rawuser
        )
        log.debug("Request headers: %s", headers)
        conn.request("GET", self.authurl, None, headers)
        res = conn.getresponse()
        reshdr = dict(res.headers)
        log.debug("Response status: %s %s", res.status, res.reason)
        log.debug("Response headers: %s", reshdr)
        log.debug("Response data: %s [...]", res.read(100))

        # Remove the reference to the socket, so that it can not be closed by
        # the response object (we want to keep the socket open)
        res.fp = None

        # Server should respond with a challenge message
        auth_header_values = reshdr[resp_header].split(", ")
        auth_header_value = None
        for s in auth_header_values:
            if s[:5] == "NTLM ":
                auth_header_value = s[5:]
        if auth_header_value is None:
            raise Exception(
                "Unexpected %s response header: %s" % (resp_header, reshdr[resp_header])
            )

        # Send authentication message
        ServerChallenge, NegotiateFlags = ntlm.parse_NTLM_CHALLENGE_MESSAGE(
            auth_header_value
        )
        auth_msg = ntlm.create_NTLM_AUTHENTICATE_MESSAGE(
            ServerChallenge, self.user, self.domain, self.pw, NegotiateFlags
        )
        headers[req_header] = "NTLM %s" % auth_msg
        log.debug("Request headers: %s", headers)
        conn.request("GET", self.authurl, None, headers)
        res = conn.getresponse()
        log.debug("Response status: %s %s", res.status, res.reason)
        log.debug("Response headers: %s", dict(res.headers))
        log.debug("Response data: %s [...]", res.read()[:100])
        if res.status != 200:
            if res.status == 401:
                raise Exception("Server rejected request: wrong username or password")
            raise Exception("Wrong server response: %s %s" % (res.status, res.reason))

        res.fp = None
        log.debug("Connection established")
        return conn

    def urlopen(
        self,
        method,
        url,
        body=None,
        headers=None,
        retries=3,
        redirect=True,
        assert_same_host=True,
    ):
        if headers is None:
            headers = {}
        headers["Connection"] = "Keep-Alive"
        return super(NTLMConnectionPool, self).urlopen(
            method, url, body, headers, retries, redirect, assert_same_host
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\__init__.py ===
"""Polynomial manipulation algorithms and algebraic objects. """

__all__ = [
    'Poly', 'PurePoly', 'poly_from_expr', 'parallel_poly_from_expr', 'degree',
    'total_degree', 'degree_list', 'LC', 'LM', 'LT', 'pdiv', 'prem', 'pquo',
    'pexquo', 'div', 'rem', 'quo', 'exquo', 'half_gcdex', 'gcdex', 'invert',
    'subresultants', 'resultant', 'discriminant', 'cofactors', 'gcd_list',
    'gcd', 'lcm_list', 'lcm', 'terms_gcd', 'trunc', 'monic', 'content',
    'primitive', 'compose', 'decompose', 'sturm', 'gff_list', 'gff',
    'sqf_norm', 'sqf_part', 'sqf_list', 'sqf', 'factor_list', 'factor',
    'intervals', 'refine_root', 'count_roots', 'all_roots', 'real_roots',
    'nroots', 'ground_roots', 'nth_power_roots_poly', 'cancel', 'reduced',
    'groebner', 'is_zero_dimensional', 'GroebnerBasis', 'poly',

    'symmetrize', 'horner', 'interpolate', 'rational_interpolate', 'viete',

    'together',

    'BasePolynomialError', 'ExactQuotientFailed', 'PolynomialDivisionFailed',
    'OperationNotSupported', 'HeuristicGCDFailed', 'HomomorphismFailed',
    'IsomorphismFailed', 'ExtraneousFactors', 'EvaluationFailed',
    'RefinementFailed', 'CoercionFailed', 'NotInvertible', 'NotReversible',
    'NotAlgebraic', 'DomainError', 'PolynomialError', 'UnificationFailed',
    'GeneratorsError', 'GeneratorsNeeded', 'ComputationFailed',
    'UnivariatePolynomialError', 'MultivariatePolynomialError',
    'PolificationFailed', 'OptionError', 'FlagError',

    'minpoly', 'minimal_polynomial', 'primitive_element', 'field_isomorphism',
    'to_number_field', 'isolate', 'round_two', 'prime_decomp',
    'prime_valuation', 'galois_group',

    'itermonomials', 'Monomial',

    'lex', 'grlex', 'grevlex', 'ilex', 'igrlex', 'igrevlex',

    'CRootOf', 'rootof', 'RootOf', 'ComplexRootOf', 'RootSum',

    'roots',

    'Domain', 'FiniteField', 'IntegerRing', 'RationalField', 'RealField',
    'ComplexField', 'PythonFiniteField', 'GMPYFiniteField',
    'PythonIntegerRing', 'GMPYIntegerRing', 'PythonRational',
    'GMPYRationalField', 'AlgebraicField', 'PolynomialRing', 'FractionField',
    'ExpressionDomain', 'FF_python', 'FF_gmpy', 'ZZ_python', 'ZZ_gmpy',
    'QQ_python', 'QQ_gmpy', 'GF', 'FF', 'ZZ', 'QQ', 'ZZ_I', 'QQ_I', 'RR',
    'CC', 'EX', 'EXRAW',

    'construct_domain',

    'swinnerton_dyer_poly', 'cyclotomic_poly', 'symmetric_poly',
    'random_poly', 'interpolating_poly',

    'jacobi_poly', 'chebyshevt_poly', 'chebyshevu_poly', 'hermite_poly',
    'hermite_prob_poly', 'legendre_poly', 'laguerre_poly',

    'bernoulli_poly', 'bernoulli_c_poly', 'genocchi_poly', 'euler_poly',
    'andre_poly',

    'apart', 'apart_list', 'assemble_partfrac_list',

    'Options',

    'ring', 'xring', 'vring', 'sring',

    'field', 'xfield', 'vfield', 'sfield'
]

from .polytools import (Poly, PurePoly, poly_from_expr,
        parallel_poly_from_expr, degree, total_degree, degree_list, LC, LM,
        LT, pdiv, prem, pquo, pexquo, div, rem, quo, exquo, half_gcdex, gcdex,
        invert, subresultants, resultant, discriminant, cofactors, gcd_list,
        gcd, lcm_list, lcm, terms_gcd, trunc, monic, content, primitive,
        compose, decompose, sturm, gff_list, gff, sqf_norm, sqf_part,
        sqf_list, sqf, factor_list, factor, intervals, refine_root,
        count_roots, all_roots, real_roots, nroots, ground_roots,
        nth_power_roots_poly, cancel, reduced, groebner, is_zero_dimensional,
        GroebnerBasis, poly)

from .polyfuncs import (symmetrize, horner, interpolate,
        rational_interpolate, viete)

from .rationaltools import together

from .polyerrors import (BasePolynomialError, ExactQuotientFailed,
        PolynomialDivisionFailed, OperationNotSupported, HeuristicGCDFailed,
        HomomorphismFailed, IsomorphismFailed, ExtraneousFactors,
        EvaluationFailed, RefinementFailed, CoercionFailed, NotInvertible,
        NotReversible, NotAlgebraic, DomainError, PolynomialError,
        UnificationFailed, GeneratorsError, GeneratorsNeeded,
        ComputationFailed, UnivariatePolynomialError,
        MultivariatePolynomialError, PolificationFailed, OptionError,
        FlagError)

from .numberfields import (minpoly, minimal_polynomial, primitive_element,
        field_isomorphism, to_number_field, isolate, round_two, prime_decomp,
        prime_valuation, galois_group)

from .monomials import itermonomials, Monomial

from .orderings import lex, grlex, grevlex, ilex, igrlex, igrevlex

from .rootoftools import CRootOf, rootof, RootOf, ComplexRootOf, RootSum

from .polyroots import roots

from .domains import (Domain, FiniteField, IntegerRing, RationalField,
        RealField, ComplexField, PythonFiniteField, GMPYFiniteField,
        PythonIntegerRing, GMPYIntegerRing, PythonRational, GMPYRationalField,
        AlgebraicField, PolynomialRing, FractionField, ExpressionDomain,
        FF_python, FF_gmpy, ZZ_python, ZZ_gmpy, QQ_python, QQ_gmpy, GF, FF,
        ZZ, QQ, ZZ_I, QQ_I, RR, CC, EX, EXRAW)

from .constructor import construct_domain

from .specialpolys import (swinnerton_dyer_poly, cyclotomic_poly,
        symmetric_poly, random_poly, interpolating_poly)

from .orthopolys import (jacobi_poly, chebyshevt_poly, chebyshevu_poly,
        hermite_poly, hermite_prob_poly, legendre_poly, laguerre_poly)

from .appellseqs import (bernoulli_poly, bernoulli_c_poly, genocchi_poly,
        euler_poly, andre_poly)

from .partfrac import apart, apart_list, assemble_partfrac_list

from .polyoptions import Options

from .rings import ring, xring, vring, sring

from .fields import field, xfield, vfield, sfield

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\google\protobuf\proto_text.py ===
# Protocol Buffers - Google's data interchange format
# Copyright 2025 Google Inc.  All rights reserved.
#
# Use of this source code is governed by a BSD-style
# license that can be found in the LICENSE file or at
# https://developers.google.com/open-source/licenses/bsd

"""Contains the Nextgen Pythonic Protobuf Text Format APIs."""
from typing import AnyStr, Callable, Optional, Text, Type, Union

from google.protobuf import text_format
from google.protobuf.descriptor_pool import DescriptorPool
from google.protobuf.message import Message

_MsgFormatter = Callable[[Message, Union[int, bool], bool], Optional[Text]]


def serialize(
    message: Message,
    as_utf8: bool = True,
    as_one_line: bool = False,
    use_short_repeated_primitives: bool = False,
    pointy_brackets: bool = False,
    use_index_order: bool = False,
    float_format: Optional[str] = None,
    double_format: Optional[str] = None,
    use_field_number: bool = False,
    descriptor_pool: Optional[DescriptorPool] = None,
    indent: int = 0,
    message_formatter: Optional[_MsgFormatter] = None,
    print_unknown_fields: bool = False,
    force_colon: bool = False,
) -> str:
  """Convert protobuf message to text format.

  Double values can be formatted compactly with 15 digits of
  precision (which is the most that IEEE 754 "double" can guarantee)
  using double_format='.15g'. To ensure that converting to text and back to a
  proto will result in an identical value, double_format='.17g' should be used.

  Args:
    message: The protocol buffers message.
    as_utf8: Return unescaped Unicode for non-ASCII characters.
    as_one_line: Don't introduce newlines between fields.
    use_short_repeated_primitives: Use short repeated format for primitives.
    pointy_brackets: If True, use angle brackets instead of curly braces for
      nesting.
    use_index_order: If True, fields of a proto message will be printed using
      the order defined in source code instead of the field number, extensions
      will be printed at the end of the message and their relative order is
      determined by the extension number. By default, use the field number
      order.
    float_format (str): If set, use this to specify float field formatting (per
      the "Format Specification Mini-Language"); otherwise, shortest float that
      has same value in wire will be printed. Also affect double field if
      double_format is not set but float_format is set.
    double_format (str): If set, use this to specify double field formatting
      (per the "Format Specification Mini-Language"); if it is not set but
      float_format is set, use float_format. Otherwise, use ``str()``
    use_field_number: If True, print field numbers instead of names.
    descriptor_pool (DescriptorPool): Descriptor pool used to resolve Any types.
    indent (int): The initial indent level, in terms of spaces, for pretty
      print.
    message_formatter (function(message, indent, as_one_line) -> unicode|None):
      Custom formatter for selected sub-messages (usually based on message
      type). Use to pretty print parts of the protobuf for easier diffing.
    print_unknown_fields: If True, unknown fields will be printed.
    force_colon: If set, a colon will be added after the field name even if the
      field is a proto message.

  Returns:
    str: A string of the text formatted protocol buffer message.
  """
  return text_format.MessageToString(
      message=message,
      as_utf8=as_utf8,
      as_one_line=as_one_line,
      use_short_repeated_primitives=use_short_repeated_primitives,
      pointy_brackets=pointy_brackets,
      use_index_order=use_index_order,
      float_format=float_format,
      double_format=double_format,
      use_field_number=use_field_number,
      descriptor_pool=descriptor_pool,
      indent=indent,
      message_formatter=message_formatter,
      print_unknown_fields=print_unknown_fields,
      force_colon=force_colon,
  )


def parse(
    message_class: Type[Message],
    text: AnyStr,
    allow_unknown_extension: bool = False,
    allow_field_number: bool = False,
    descriptor_pool: Optional[DescriptorPool] = None,
    allow_unknown_field: bool = False,
) -> Message:
  """Parses a text representation of a protocol message into a message.

  Args:
    message_class: The message meta class.
    text (str): Message text representation.
    message (Message): A protocol buffer message to merge into.
    allow_unknown_extension: if True, skip over missing extensions and keep
      parsing
    allow_field_number: if True, both field number and field name are allowed.
    descriptor_pool (DescriptorPool): Descriptor pool used to resolve Any types.
    allow_unknown_field: if True, skip over unknown field and keep parsing.
      Avoid to use this option if possible. It may hide some errors (e.g.
      spelling error on field name)

  Returns:
    Message: A new message passed from text.

  Raises:
    ParseError: On text parsing problems.
  """
  new_message = message_class()
  text_format.Parse(
      text=text,
      message=new_message,
      allow_unknown_extension=allow_unknown_extension,
      allow_field_number=allow_field_number,
      descriptor_pool=descriptor_pool,
      allow_unknown_field=allow_unknown_field,
  )
  return new_message

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\tools\pytorch_export_contrib_ops.py ===
# Copyright (c) Microsoft Corporation. All rights reserved.
# Licensed under the MIT License.

"""
Support for registering ONNX Runtime's built-in contrib ops with
PyTorch-ONNX exporter (torch.onnx.export).
"""

import typing

try:
    # TODO(justinchuby): Create a function to alert users when torch is not installed
    import torch
except ModuleNotFoundError:
    raise ModuleNotFoundError(  # noqa: B904
        "This module is only useful in combination with PyTorch. To install PyTorch see https://pytorch.org/."
    )

from torch.onnx import symbolic_helper

_OPSET_VERSION = 1
_registered_ops: typing.AbstractSet[str] = set()


def _reg(symbolic_fn: typing.Callable, namespace: str = ""):
    name = f"{namespace}::{symbolic_fn.__name__}"
    torch.onnx.register_custom_op_symbolic(name, symbolic_fn, _OPSET_VERSION)
    _registered_ops.add(name)


def register():
    """Register ONNX Runtime's built-in contrib ops.

    Should be run before torch.onnx.export().
    """

    def grid_sampler(g, input, grid, mode, padding_mode, align_corners):
        # mode
        #   'bilinear'      : onnx::Constant[value={0}]
        #   'nearest'       : onnx::Constant[value={1}]
        #   'bicubic'       : onnx::Constant[value={2}]
        # padding_mode
        #   'zeros'         : onnx::Constant[value={0}]
        #   'border'        : onnx::Constant[value={1}]
        #   'reflection'    : onnx::Constant[value={2}]
        mode = symbolic_helper._maybe_get_const(mode, "i")
        padding_mode = symbolic_helper._maybe_get_const(padding_mode, "i")
        mode_str = ["bilinear", "nearest", "bicubic"][mode]
        padding_mode_str = ["zeros", "border", "reflection"][padding_mode]
        align_corners = int(symbolic_helper._maybe_get_const(align_corners, "b"))

        # From opset v13 onward, the output shape can be specified with
        # (N, C, H, W) (N, H_out, W_out, 2) => (N, C, H_out, W_out)
        # input_shape = input.type().sizes()
        # gird_shape = grid.type().sizes()
        # output_shape = input_shape[:2] + gird_shape[1:3]
        # g.op(...).setType(input.type().with_sizes(output_shape))

        return g.op(
            "com.microsoft::GridSample",
            input,
            grid,
            mode_s=mode_str,
            padding_mode_s=padding_mode_str,
            align_corners_i=align_corners,
        )

    _reg(grid_sampler)

    def inverse(g, self):
        return g.op("com.microsoft::Inverse", self).setType(self.type())

    _reg(inverse)

    @torch.onnx.symbolic_helper.parse_args("v", "s")
    def gelu(g, self: torch._C.Value, approximate: str = "none"):
        # Use microsoft::Gelu for performance if possible. It only supports approximate == "none"
        if approximate == "none":
            return g.op("com.microsoft::Gelu", self).setType(self.type())
        return torch.onnx.symbolic_opset9.gelu(g, self, approximate)

    _reg(gelu)

    def triu(g, self, diagonal):
        return g.op("com.microsoft::Trilu", self, diagonal, upper_i=1).setType(self.type())

    _reg(triu)

    def tril(g, self, diagonal):
        return g.op("com.microsoft::Trilu", self, diagonal, upper_i=0).setType(self.type())

    _reg(tril)

    @torch.onnx.symbolic_helper.parse_args("v")
    def DynamicTimeWarping(g, self):  # noqa: N802
        return g.op("com.microsoft::DynamicTimeWarping", self)

    _reg(DynamicTimeWarping, namespace="onnxruntime")

    def UnfoldTensor(g, self, dim, size, step):  # noqa: N802
        dim = int(symbolic_helper._maybe_get_const(dim, "i"))
        size = int(symbolic_helper._maybe_get_const(size, "i"))
        step = int(symbolic_helper._maybe_get_const(step, "i"))
        return g.op(
            "com.microsoft::UnfoldTensor",
            self,
            dim_i=dim,
            size_i=size,
            step_i=step,
        ).setType(self.type().with_sizes([None, None, None, None, size]))

    _reg(UnfoldTensor, namespace="onnxruntime")


def unregister():
    """Unregister ONNX Runtime's built-in contrib ops."""
    for name in _registered_ops:
        try:
            torch.onnx.unregister_custom_op_symbolic(name, _OPSET_VERSION)
        except AttributeError:
            # The symbolic_registry module was removed in PyTorch 1.13.
            # We are importing it here for backwards compatibility
            # because unregister_custom_op_symbolic is not available before PyTorch 1.12
            from torch.onnx import symbolic_registry

            namespace, kind = name.split("::")
            for version in symbolic_helper._onnx_stable_opsets:
                if version >= _OPSET_VERSION and symbolic_registry.is_registered_op(kind, namespace, version):
                    del symbolic_registry._registry[(namespace, version)][kind]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\chart\line_chart.py ===
#Autogenerated schema
from openpyxl.descriptors import (
    Typed,
    Sequence,
    Alias,
    )
from openpyxl.descriptors.excel import ExtensionList
from openpyxl.descriptors.nested import (
    NestedSet,
    NestedBool,
)

from ._chart import ChartBase
from .updown_bars import UpDownBars
from .descriptors import NestedGapAmount
from .axis import TextAxis, NumericAxis, SeriesAxis, ChartLines, _BaseAxis
from .label import DataLabelList
from .series import Series


class _LineChartBase(ChartBase):

    grouping = NestedSet(values=(['percentStacked', 'standard', 'stacked']))
    varyColors = NestedBool(allow_none=True)
    ser = Sequence(expected_type=Series, allow_none=True)
    dLbls = Typed(expected_type=DataLabelList, allow_none=True)
    dataLabels = Alias("dLbls")
    dropLines = Typed(expected_type=ChartLines, allow_none=True)

    _series_type = "line"

    __elements__ = ('grouping', 'varyColors', 'ser', 'dLbls', 'dropLines')

    def __init__(self,
                 grouping="standard",
                 varyColors=None,
                 ser=(),
                 dLbls=None,
                 dropLines=None,
                 **kw
                ):
        self.grouping = grouping
        self.varyColors = varyColors
        self.ser = ser
        self.dLbls = dLbls
        self.dropLines = dropLines
        super().__init__(**kw)


class LineChart(_LineChartBase):

    tagname = "lineChart"

    grouping = _LineChartBase.grouping
    varyColors = _LineChartBase.varyColors
    ser = _LineChartBase.ser
    dLbls = _LineChartBase.dLbls
    dropLines =_LineChartBase.dropLines

    hiLowLines = Typed(expected_type=ChartLines, allow_none=True)
    upDownBars = Typed(expected_type=UpDownBars, allow_none=True)
    marker = NestedBool(allow_none=True)
    smooth = NestedBool(allow_none=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    x_axis = Typed(expected_type=_BaseAxis)
    y_axis = Typed(expected_type=NumericAxis)

    __elements__ = _LineChartBase.__elements__ + ('hiLowLines', 'upDownBars', 'marker', 'smooth', 'axId')

    def __init__(self,
                 hiLowLines=None,
                 upDownBars=None,
                 marker=None,
                 smooth=None,
                 extLst=None,
                 **kw
                ):
        self.hiLowLines = hiLowLines
        self.upDownBars = upDownBars
        self.marker = marker
        self.smooth = smooth
        self.x_axis = TextAxis()
        self.y_axis = NumericAxis()

        super().__init__(**kw)


class LineChart3D(_LineChartBase):

    tagname = "line3DChart"

    grouping = _LineChartBase.grouping
    varyColors = _LineChartBase.varyColors
    ser = _LineChartBase.ser
    dLbls = _LineChartBase.dLbls
    dropLines =_LineChartBase.dropLines

    gapDepth = NestedGapAmount()
    hiLowLines = Typed(expected_type=ChartLines, allow_none=True)
    upDownBars = Typed(expected_type=UpDownBars, allow_none=True)
    marker = NestedBool(allow_none=True)
    smooth = NestedBool(allow_none=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    x_axis = Typed(expected_type=TextAxis)
    y_axis = Typed(expected_type=NumericAxis)
    z_axis = Typed(expected_type=SeriesAxis)

    __elements__ = _LineChartBase.__elements__ + ('gapDepth', 'hiLowLines',
                                                  'upDownBars', 'marker', 'smooth', 'axId')

    def __init__(self,
                 gapDepth=None,
                 hiLowLines=None,
                 upDownBars=None,
                 marker=None,
                 smooth=None,
                 **kw
                ):
        self.gapDepth = gapDepth
        self.hiLowLines = hiLowLines
        self.upDownBars = upDownBars
        self.marker = marker
        self.smooth = smooth
        self.x_axis = TextAxis()
        self.y_axis = NumericAxis()
        self.z_axis = SeriesAxis()
        super(LineChart3D, self).__init__(**kw)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\descriptors\nested.py ===
# Copyright (c) 2010-2024 openpyxl

"""
Generic serialisable classes
"""
from .base import (
    Convertible,
    Bool,
    Descriptor,
    NoneSet,
    MinMax,
    Set,
    Float,
    Integer,
    String,
    )
from openpyxl.compat import safe_string
from openpyxl.xml.functions import Element, localname, whitespace


class Nested(Descriptor):

    nested = True
    attribute = "val"

    def __set__(self, instance, value):
        if hasattr(value, "tag"):
            tag = localname(value)
            if tag != self.name:
                raise ValueError("Tag does not match attribute")

            value = self.from_tree(value)
        super().__set__(instance, value)


    def from_tree(self, node):
        return node.get(self.attribute)


    def to_tree(self, tagname=None, value=None, namespace=None):
        namespace = getattr(self, "namespace", namespace)
        if value is not None:
            if namespace is not None:
                tagname = "{%s}%s" % (namespace, tagname)
            value = safe_string(value)
            return Element(tagname, {self.attribute:value})


class NestedValue(Nested, Convertible):
    """
    Nested tag storing the value on the 'val' attribute
    """
    pass


class NestedText(NestedValue):
    """
    Represents any nested tag with the value as the contents of the tag
    """


    def from_tree(self, node):
        return node.text


    def to_tree(self, tagname=None, value=None, namespace=None):
        namespace = getattr(self, "namespace", namespace)
        if value is not None:
            if namespace is not None:
                tagname = "{%s}%s" % (namespace, tagname)
            el = Element(tagname)
            el.text = safe_string(value)
            whitespace(el)
            return el


class NestedFloat(NestedValue, Float):

    pass


class NestedInteger(NestedValue, Integer):

    pass


class NestedString(NestedValue, String):

    pass


class NestedBool(NestedValue, Bool):


    def from_tree(self, node):
        return node.get("val", True)


class NestedNoneSet(Nested, NoneSet):

    pass


class NestedSet(Nested, Set):

    pass


class NestedMinMax(Nested, MinMax):

    pass


class EmptyTag(Nested, Bool):

    """
    Boolean if a tag exists or not.
    """

    def from_tree(self, node):
        return True


    def to_tree(self, tagname=None, value=None, namespace=None):
        if value:
            namespace = getattr(self, "namespace", namespace)
            if namespace is not None:
                tagname = "{%s}%s" % (namespace, tagname)
            return Element(tagname)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\xml\constants.py ===
# Copyright (c) 2010-2024 openpyxl


"""Constants for fixed paths in a file and xml namespace urls."""

MIN_ROW = 0
MIN_COLUMN = 0
MAX_COLUMN = 16384
MAX_ROW = 1048576

# constants
PACKAGE_PROPS = 'docProps'
PACKAGE_XL = 'xl'
PACKAGE_RELS = '_rels'
PACKAGE_THEME = PACKAGE_XL + '/' + 'theme'
PACKAGE_WORKSHEETS = PACKAGE_XL + '/' + 'worksheets'
PACKAGE_CHARTSHEETS = PACKAGE_XL + '/' + 'chartsheets'
PACKAGE_DRAWINGS = PACKAGE_XL + '/' + 'drawings'
PACKAGE_CHARTS = PACKAGE_XL + '/' + 'charts'
PACKAGE_IMAGES = PACKAGE_XL + '/' + 'media'
PACKAGE_WORKSHEET_RELS = PACKAGE_WORKSHEETS + '/' + '_rels'
PACKAGE_CHARTSHEETS_RELS = PACKAGE_CHARTSHEETS + '/' + '_rels'
PACKAGE_PIVOT_TABLE = PACKAGE_XL + '/' + 'pivotTables'
PACKAGE_PIVOT_CACHE = PACKAGE_XL + '/' + 'pivotCache'

ARC_CONTENT_TYPES = '[Content_Types].xml'
ARC_ROOT_RELS = PACKAGE_RELS + '/.rels'
ARC_WORKBOOK_RELS = PACKAGE_XL + '/' + PACKAGE_RELS + '/workbook.xml.rels'
ARC_CORE = PACKAGE_PROPS + '/core.xml'
ARC_APP = PACKAGE_PROPS + '/app.xml'
ARC_CUSTOM = PACKAGE_PROPS + '/custom.xml'
ARC_WORKBOOK = PACKAGE_XL + '/workbook.xml'
ARC_STYLE = PACKAGE_XL + '/styles.xml'
ARC_THEME = PACKAGE_THEME + '/theme1.xml'
ARC_SHARED_STRINGS = PACKAGE_XL + '/sharedStrings.xml'
ARC_CUSTOM_UI = 'customUI/customUI.xml'

## namespaces
# XML
XML_NS = "http://www.w3.org/XML/1998/namespace"
# Dublin Core
DCORE_NS = 'http://purl.org/dc/elements/1.1/'
DCTERMS_NS = 'http://purl.org/dc/terms/'
DCTERMS_PREFIX = 'dcterms'

# Document
DOC_NS = "http://schemas.openxmlformats.org/officeDocument/2006/"
REL_NS = DOC_NS + "relationships"
COMMENTS_NS = REL_NS + "/comments"
IMAGE_NS = REL_NS + "/image"
VML_NS =  REL_NS + "/vmlDrawing"
VTYPES_NS = DOC_NS + 'docPropsVTypes'
XPROPS_NS = DOC_NS + 'extended-properties'
CUSTPROPS_NS = DOC_NS + 'custom-properties'
EXTERNAL_LINK_NS = REL_NS + "/externalLink"

# CustomDocumentProperty FMTID:
CPROPS_FMTID = "{D5CDD505-2E9C-101B-9397-08002B2CF9AE}"

# Package
PKG_NS = "http://schemas.openxmlformats.org/package/2006/"
PKG_REL_NS = PKG_NS + "relationships"
COREPROPS_NS = PKG_NS + 'metadata/core-properties'
CONTYPES_NS = PKG_NS + 'content-types'

XSI_NS = 'http://www.w3.org/2001/XMLSchema-instance'
XML_NS = 'http://www.w3.org/XML/1998/namespace'
SHEET_MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

# Drawing
CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
SHEET_DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
CHART_DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/chartDrawing"

CUSTOMUI_NS = 'http://schemas.microsoft.com/office/2006/relationships/ui/extensibility'


NAMESPACES = {
    'cp': COREPROPS_NS,
    'dc': DCORE_NS,
    DCTERMS_PREFIX: DCTERMS_NS,
    'dcmitype': 'http://purl.org/dc/dcmitype/',
    'xsi': XSI_NS,
    'vt': VTYPES_NS,
    'xml': XML_NS,
    'main': SHEET_MAIN_NS,
    'cust': CUSTPROPS_NS,
}

## Mime types
WORKBOOK_MACRO = "application/vnd.ms-excel.%s.macroEnabled.main+xml"
WORKBOOK = "application/vnd.openxmlformats-officedocument.spreadsheetml.%s.main+xml"
SPREADSHEET = "application/vnd.openxmlformats-officedocument.spreadsheetml.%s+xml"
SHARED_STRINGS = SPREADSHEET % "sharedStrings"
EXTERNAL_LINK = SPREADSHEET % "externalLink"
WORKSHEET_TYPE = SPREADSHEET % "worksheet"
COMMENTS_TYPE = SPREADSHEET % "comments"
STYLES_TYPE = SPREADSHEET % "styles"
CHARTSHEET_TYPE = SPREADSHEET % "chartsheet"
DRAWING_TYPE = "application/vnd.openxmlformats-officedocument.drawing+xml"
CHART_TYPE = "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"
CHARTSHAPE_TYPE = "application/vnd.openxmlformats-officedocument.drawingml.chartshapes+xml"
THEME_TYPE = "application/vnd.openxmlformats-officedocument.theme+xml"
CPROPS_TYPE = "application/vnd.openxmlformats-officedocument.custom-properties+xml"
XLTM = WORKBOOK_MACRO % 'template'
XLSM = WORKBOOK_MACRO % 'sheet'
XLTX = WORKBOOK % 'template'
XLSX = WORKBOOK % 'sheet'


# Extensions to the specification

EXT_TYPES = {
    '{78C0D931-6437-407D-A8EE-F0AAD7539E65}': 'Conditional Formatting',
    '{CCE6A557-97BC-4B89-ADB6-D9C93CAAB3DF}': 'Data Validation',
    '{05C60535-1F16-4FD2-B633-F4F36F0B64E0}': 'Sparkline Group',
    '{A8765BA9-456A-4DAB-B4F3-ACF838C121DE}': 'Slicer List',
    '{FC87AEE6-9EDD-4A0A-B7FB-166176984837}': 'Protected Range',
    '{01252117-D84E-4E92-8308-4BE1C098FCBB}': 'Ignored Error',
    '{F7C9EE02-42E1-4005-9D12-6889AFFD525C}': 'Web Extension',
    '{3A4CF648-6AED-40f4-86FF-DC5316D8AED3}': 'Slicer List',
    '{7E03D99C-DC04-49d9-9315-930204A7B6E9}': 'Timeline Ref',
}

# Objects related to macros that we preserve
CTRL = "application/vnd.ms-excel.controlproperties+xml"
ACTIVEX = "application/vnd.ms-office.activeX+xml"
VBA = "application/vnd.ms-office.vbaProject"

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\pygments\formatter.py ===
"""
    pygments.formatter
    ~~~~~~~~~~~~~~~~~~

    Base formatter class.

    :copyright: Copyright 2006-2025 by the Pygments team, see AUTHORS.
    :license: BSD, see LICENSE for details.
"""

import codecs

from pip._vendor.pygments.util import get_bool_opt
from pip._vendor.pygments.styles import get_style_by_name

__all__ = ['Formatter']


def _lookup_style(style):
    if isinstance(style, str):
        return get_style_by_name(style)
    return style


class Formatter:
    """
    Converts a token stream to text.

    Formatters should have attributes to help selecting them. These
    are similar to the corresponding :class:`~pygments.lexer.Lexer`
    attributes.

    .. autoattribute:: name
       :no-value:

    .. autoattribute:: aliases
       :no-value:

    .. autoattribute:: filenames
       :no-value:

    You can pass options as keyword arguments to the constructor.
    All formatters accept these basic options:

    ``style``
        The style to use, can be a string or a Style subclass
        (default: "default"). Not used by e.g. the
        TerminalFormatter.
    ``full``
        Tells the formatter to output a "full" document, i.e.
        a complete self-contained document. This doesn't have
        any effect for some formatters (default: false).
    ``title``
        If ``full`` is true, the title that should be used to
        caption the document (default: '').
    ``encoding``
        If given, must be an encoding name. This will be used to
        convert the Unicode token strings to byte strings in the
        output. If it is "" or None, Unicode strings will be written
        to the output file, which most file-like objects do not
        support (default: None).
    ``outencoding``
        Overrides ``encoding`` if given.

    """

    #: Full name for the formatter, in human-readable form.
    name = None

    #: A list of short, unique identifiers that can be used to lookup
    #: the formatter from a list, e.g. using :func:`.get_formatter_by_name()`.
    aliases = []

    #: A list of fnmatch patterns that match filenames for which this
    #: formatter can produce output. The patterns in this list should be unique
    #: among all formatters.
    filenames = []

    #: If True, this formatter outputs Unicode strings when no encoding
    #: option is given.
    unicodeoutput = True

    def __init__(self, **options):
        """
        As with lexers, this constructor takes arbitrary optional arguments,
        and if you override it, you should first process your own options, then
        call the base class implementation.
        """
        self.style = _lookup_style(options.get('style', 'default'))
        self.full = get_bool_opt(options, 'full', False)
        self.title = options.get('title', '')
        self.encoding = options.get('encoding', None) or None
        if self.encoding in ('guess', 'chardet'):
            # can happen for e.g. pygmentize -O encoding=guess
            self.encoding = 'utf-8'
        self.encoding = options.get('outencoding') or self.encoding
        self.options = options

    def get_style_defs(self, arg=''):
        """
        This method must return statements or declarations suitable to define
        the current style for subsequent highlighted text (e.g. CSS classes
        in the `HTMLFormatter`).

        The optional argument `arg` can be used to modify the generation and
        is formatter dependent (it is standardized because it can be given on
        the command line).

        This method is called by the ``-S`` :doc:`command-line option <cmdline>`,
        the `arg` is then given by the ``-a`` option.
        """
        return ''

    def format(self, tokensource, outfile):
        """
        This method must format the tokens from the `tokensource` iterable and
        write the formatted version to the file object `outfile`.

        Formatter options can control how exactly the tokens are converted.
        """
        if self.encoding:
            # wrap the outfile in a StreamWriter
            outfile = codecs.lookup(self.encoding)[3](outfile)
        return self.format_unencoded(tokensource, outfile)

    # Allow writing Formatter[str] or Formatter[bytes]. That's equivalent to
    # Formatter. This helps when using third-party type stubs from typeshed.
    def __class_getitem__(cls, name):
        return cls

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\mssql\json.py ===
# dialects/mssql/json.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors

from ... import types as sqltypes

# technically, all the dialect-specific datatypes that don't have any special
# behaviors would be private with names like _MSJson. However, we haven't been
# doing this for mysql.JSON or sqlite.JSON which both have JSON / JSONIndexType
# / JSONPathType in their json.py files, so keep consistent with that
# sub-convention for now.  A future change can update them all to be
# package-private at once.


class JSON(sqltypes.JSON):
    """MSSQL JSON type.

    MSSQL supports JSON-formatted data as of SQL Server 2016.

    The :class:`_mssql.JSON` datatype at the DDL level will represent the
    datatype as ``NVARCHAR(max)``, but provides for JSON-level comparison
    functions as well as Python coercion behavior.

    :class:`_mssql.JSON` is used automatically whenever the base
    :class:`_types.JSON` datatype is used against a SQL Server backend.

    .. seealso::

        :class:`_types.JSON` - main documentation for the generic
        cross-platform JSON datatype.

    The :class:`_mssql.JSON` type supports persistence of JSON values
    as well as the core index operations provided by :class:`_types.JSON`
    datatype, by adapting the operations to render the ``JSON_VALUE``
    or ``JSON_QUERY`` functions at the database level.

    The SQL Server :class:`_mssql.JSON` type necessarily makes use of the
    ``JSON_QUERY`` and ``JSON_VALUE`` functions when querying for elements
    of a JSON object.   These two functions have a major restriction in that
    they are **mutually exclusive** based on the type of object to be returned.
    The ``JSON_QUERY`` function **only** returns a JSON dictionary or list,
    but not an individual string, numeric, or boolean element; the
    ``JSON_VALUE`` function **only** returns an individual string, numeric,
    or boolean element.   **both functions either return NULL or raise
    an error if they are not used against the correct expected value**.

    To handle this awkward requirement, indexed access rules are as follows:

    1. When extracting a sub element from a JSON that is itself a JSON
       dictionary or list, the :meth:`_types.JSON.Comparator.as_json` accessor
       should be used::

            stmt = select(data_table.c.data["some key"].as_json()).where(
                data_table.c.data["some key"].as_json() == {"sub": "structure"}
            )

    2. When extracting a sub element from a JSON that is a plain boolean,
       string, integer, or float, use the appropriate method among
       :meth:`_types.JSON.Comparator.as_boolean`,
       :meth:`_types.JSON.Comparator.as_string`,
       :meth:`_types.JSON.Comparator.as_integer`,
       :meth:`_types.JSON.Comparator.as_float`::

            stmt = select(data_table.c.data["some key"].as_string()).where(
                data_table.c.data["some key"].as_string() == "some string"
            )

    .. versionadded:: 1.4


    """

    # note there was a result processor here that was looking for "number",
    # but none of the tests seem to exercise it.


# Note: these objects currently match exactly those of MySQL, however since
# these are not generalizable to all JSON implementations, remain separately
# implemented for each dialect.
class _FormatTypeMixin:
    def _format_value(self, value):
        raise NotImplementedError()

    def bind_processor(self, dialect):
        super_proc = self.string_bind_processor(dialect)

        def process(value):
            value = self._format_value(value)
            if super_proc:
                value = super_proc(value)
            return value

        return process

    def literal_processor(self, dialect):
        super_proc = self.string_literal_processor(dialect)

        def process(value):
            value = self._format_value(value)
            if super_proc:
                value = super_proc(value)
            return value

        return process


class JSONIndexType(_FormatTypeMixin, sqltypes.JSON.JSONIndexType):
    def _format_value(self, value):
        if isinstance(value, int):
            value = "$[%s]" % value
        else:
            value = '$."%s"' % value
        return value


class JSONPathType(_FormatTypeMixin, sqltypes.JSON.JSONPathType):
    def _format_value(self, value):
        return "$%s" % (
            "".join(
                [
                    "[%s]" % elem if isinstance(elem, int) else '."%s"' % elem
                    for elem in value
                ]
            )
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\postgresql\operators.py ===
# dialects/postgresql/operators.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors
from ...sql import operators


_getitem_precedence = operators._PRECEDENCE[operators.json_getitem_op]
_eq_precedence = operators._PRECEDENCE[operators.eq]

# JSON + JSONB
ASTEXT = operators.custom_op(
    "->>",
    precedence=_getitem_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
)

JSONPATH_ASTEXT = operators.custom_op(
    "#>>",
    precedence=_getitem_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
)

# JSONB + HSTORE
HAS_KEY = operators.custom_op(
    "?",
    precedence=_eq_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
    is_comparison=True,
)

HAS_ALL = operators.custom_op(
    "?&",
    precedence=_eq_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
    is_comparison=True,
)

HAS_ANY = operators.custom_op(
    "?|",
    precedence=_eq_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
    is_comparison=True,
)

# JSONB
DELETE_PATH = operators.custom_op(
    "#-",
    precedence=_getitem_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
)

PATH_EXISTS = operators.custom_op(
    "@?",
    precedence=_eq_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
    is_comparison=True,
)

PATH_MATCH = operators.custom_op(
    "@@",
    precedence=_eq_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
    is_comparison=True,
)

# JSONB + ARRAY + HSTORE + RANGE
CONTAINS = operators.custom_op(
    "@>",
    precedence=_eq_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
    is_comparison=True,
)

CONTAINED_BY = operators.custom_op(
    "<@",
    precedence=_eq_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
    is_comparison=True,
)

# ARRAY + RANGE
OVERLAP = operators.custom_op(
    "&&",
    precedence=_eq_precedence,
    is_comparison=True,
)

# RANGE
STRICTLY_LEFT_OF = operators.custom_op(
    "<<", precedence=_eq_precedence, is_comparison=True
)

STRICTLY_RIGHT_OF = operators.custom_op(
    ">>", precedence=_eq_precedence, is_comparison=True
)

NOT_EXTEND_RIGHT_OF = operators.custom_op(
    "&<", precedence=_eq_precedence, is_comparison=True
)

NOT_EXTEND_LEFT_OF = operators.custom_op(
    "&>", precedence=_eq_precedence, is_comparison=True
)

ADJACENT_TO = operators.custom_op(
    "-|-", precedence=_eq_precedence, is_comparison=True
)

# HSTORE
GETITEM = operators.custom_op(
    "->",
    precedence=_getitem_precedence,
    natural_self_precedent=True,
    eager_grouping=True,
)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\tensor\array\array_derivatives.py ===
from __future__ import annotations

from sympy.core.expr import Expr
from sympy.core.function import Derivative
from sympy.core.numbers import Integer
from sympy.matrices.matrixbase import MatrixBase
from .ndim_array import NDimArray
from .arrayop import derive_by_array
from sympy.matrices.expressions.matexpr import MatrixExpr
from sympy.matrices.expressions.special import ZeroMatrix
from sympy.matrices.expressions.matexpr import _matrix_derivative


class ArrayDerivative(Derivative):

    is_scalar = False

    def __new__(cls, expr, *variables, **kwargs):
        obj = super().__new__(cls, expr, *variables, **kwargs)
        if isinstance(obj, ArrayDerivative):
            obj._shape = obj._get_shape()
        return obj

    def _get_shape(self):
        shape = ()
        for v, count in self.variable_count:
            if hasattr(v, "shape"):
                for i in range(count):
                    shape += v.shape
        if hasattr(self.expr, "shape"):
            shape += self.expr.shape
        return shape

    @property
    def shape(self):
        return self._shape

    @classmethod
    def _get_zero_with_shape_like(cls, expr):
        if isinstance(expr, (MatrixBase, NDimArray)):
            return expr.zeros(*expr.shape)
        elif isinstance(expr, MatrixExpr):
            return ZeroMatrix(*expr.shape)
        else:
            raise RuntimeError("Unable to determine shape of array-derivative.")

    @staticmethod
    def _call_derive_scalar_by_matrix(expr: Expr, v: MatrixBase) -> Expr:
        return v.applyfunc(lambda x: expr.diff(x))

    @staticmethod
    def _call_derive_scalar_by_matexpr(expr: Expr, v: MatrixExpr) -> Expr:
        if expr.has(v):
            return _matrix_derivative(expr, v)
        else:
            return ZeroMatrix(*v.shape)

    @staticmethod
    def _call_derive_scalar_by_array(expr: Expr, v: NDimArray) -> Expr:
        return v.applyfunc(lambda x: expr.diff(x))

    @staticmethod
    def _call_derive_matrix_by_scalar(expr: MatrixBase, v: Expr) -> Expr:
        return _matrix_derivative(expr, v)

    @staticmethod
    def _call_derive_matexpr_by_scalar(expr: MatrixExpr, v: Expr) -> Expr:
        return expr._eval_derivative(v)

    @staticmethod
    def _call_derive_array_by_scalar(expr: NDimArray, v: Expr) -> Expr:
        return expr.applyfunc(lambda x: x.diff(v))

    @staticmethod
    def _call_derive_default(expr: Expr, v: Expr) -> Expr | None:
        if expr.has(v):
            return _matrix_derivative(expr, v)
        else:
            return None

    @classmethod
    def _dispatch_eval_derivative_n_times(cls, expr, v, count):
        # Evaluate the derivative `n` times.  If
        # `_eval_derivative_n_times` is not overridden by the current
        # object, the default in `Basic` will call a loop over
        # `_eval_derivative`:

        if not isinstance(count, (int, Integer)) or ((count <= 0) == True):
            return None

        # TODO: this could be done with multiple-dispatching:
        if expr.is_scalar:
            if isinstance(v, MatrixBase):
                result = cls._call_derive_scalar_by_matrix(expr, v)
            elif isinstance(v, MatrixExpr):
                result = cls._call_derive_scalar_by_matexpr(expr, v)
            elif isinstance(v, NDimArray):
                result = cls._call_derive_scalar_by_array(expr, v)
            elif v.is_scalar:
                # scalar by scalar has a special
                return super()._dispatch_eval_derivative_n_times(expr, v, count)
            else:
                return None
        elif v.is_scalar:
            if isinstance(expr, MatrixBase):
                result = cls._call_derive_matrix_by_scalar(expr, v)
            elif isinstance(expr, MatrixExpr):
                result = cls._call_derive_matexpr_by_scalar(expr, v)
            elif isinstance(expr, NDimArray):
                result = cls._call_derive_array_by_scalar(expr, v)
            else:
                return None
        else:
            # Both `expr` and `v` are some array/matrix type:
            if isinstance(expr, MatrixBase) or isinstance(v, MatrixBase):
                result = derive_by_array(expr, v)
            elif isinstance(expr, MatrixExpr) and isinstance(v, MatrixExpr):
                result = cls._call_derive_default(expr, v)
            elif isinstance(expr, MatrixExpr) or isinstance(v, MatrixExpr):
                # if one expression is a symbolic matrix expression while the other isn't, don't evaluate:
                return None
            else:
                result = derive_by_array(expr, v)
        if result is None:
            return None
        if count == 1:
            return result
        else:
            return cls._dispatch_eval_derivative_n_times(result, v, count - 1)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_highlevel_generic.py ===
from __future__ import annotations

from typing import TYPE_CHECKING, Generic, TypeVar

import attrs

import trio
from trio._util import final

from .abc import AsyncResource, HalfCloseableStream, ReceiveStream, SendStream

if TYPE_CHECKING:
    from typing_extensions import TypeGuard


SendStreamT = TypeVar("SendStreamT", bound=SendStream)
ReceiveStreamT = TypeVar("ReceiveStreamT", bound=ReceiveStream)


async def aclose_forcefully(resource: AsyncResource) -> None:
    """Close an async resource or async generator immediately, without
    blocking to do any graceful cleanup.

    :class:`~trio.abc.AsyncResource` objects guarantee that if their
    :meth:`~trio.abc.AsyncResource.aclose` method is cancelled, then they will
    still close the resource (albeit in a potentially ungraceful
    fashion). :func:`aclose_forcefully` is a convenience function that
    exploits this behavior to let you force a resource to be closed without
    blocking: it works by calling ``await resource.aclose()`` and then
    cancelling it immediately.

    Most users won't need this, but it may be useful on cleanup paths where
    you can't afford to block, or if you want to close a resource and don't
    care about handling it gracefully. For example, if
    :class:`~trio.SSLStream` encounters an error and cannot perform its
    own graceful close, then there's no point in waiting to gracefully shut
    down the underlying transport either, so it calls ``await
    aclose_forcefully(self.transport_stream)``.

    Note that this function is async, and that it acts as a checkpoint, but
    unlike most async functions it cannot block indefinitely (at least,
    assuming the underlying resource object is correctly implemented).

    """
    with trio.CancelScope() as cs:
        cs.cancel()
        await resource.aclose()


def _is_halfclosable(stream: SendStream) -> TypeGuard[HalfCloseableStream]:
    """Check if the stream has a send_eof() method."""
    return hasattr(stream, "send_eof")


@final
@attrs.define(eq=False, slots=False)
class StapledStream(
    HalfCloseableStream,
    Generic[SendStreamT, ReceiveStreamT],
):
    """This class `staples <https://en.wikipedia.org/wiki/Staple_(fastener)>`__
    together two unidirectional streams to make single bidirectional stream.

    Args:
      send_stream (~trio.abc.SendStream): The stream to use for sending.
      receive_stream (~trio.abc.ReceiveStream): The stream to use for
          receiving.

    Example:

       A silly way to make a stream that echoes back whatever you write to
       it::

          left, right = trio.testing.memory_stream_pair()
          echo_stream = StapledStream(SocketStream(left), SocketStream(right))
          await echo_stream.send_all(b"x")
          assert await echo_stream.receive_some() == b"x"

    :class:`StapledStream` objects implement the methods in the
    :class:`~trio.abc.HalfCloseableStream` interface. They also have two
    additional public attributes:

    .. attribute:: send_stream

       The underlying :class:`~trio.abc.SendStream`. :meth:`send_all` and
       :meth:`wait_send_all_might_not_block` are delegated to this object.

    .. attribute:: receive_stream

       The underlying :class:`~trio.abc.ReceiveStream`. :meth:`receive_some`
       is delegated to this object.

    """

    send_stream: SendStreamT
    receive_stream: ReceiveStreamT

    async def send_all(self, data: bytes | bytearray | memoryview) -> None:
        """Calls ``self.send_stream.send_all``."""
        return await self.send_stream.send_all(data)

    async def wait_send_all_might_not_block(self) -> None:
        """Calls ``self.send_stream.wait_send_all_might_not_block``."""
        return await self.send_stream.wait_send_all_might_not_block()

    async def send_eof(self) -> None:
        """Shuts down the send side of the stream.

        If :meth:`self.send_stream.send_eof() <trio.abc.HalfCloseableStream.send_eof>` exists,
        then this calls it. Otherwise, this calls
        :meth:`self.send_stream.aclose() <trio.abc.AsyncResource.aclose>`.
        """
        stream = self.send_stream
        if _is_halfclosable(stream):
            return await stream.send_eof()
        else:
            return await stream.aclose()

    # we intentionally accept more types from the caller than we support returning
    async def receive_some(self, max_bytes: int | None = None) -> bytes:
        """Calls ``self.receive_stream.receive_some``."""
        return await self.receive_stream.receive_some(max_bytes)

    async def aclose(self) -> None:
        """Calls ``aclose`` on both underlying streams."""
        try:
            await self.send_stream.aclose()
        finally:
            await self.receive_stream.aclose()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\flask\blueprints.py ===
from __future__ import annotations

import os
import typing as t
from datetime import timedelta

from .cli import AppGroup
from .globals import current_app
from .helpers import send_from_directory
from .sansio.blueprints import Blueprint as SansioBlueprint
from .sansio.blueprints import BlueprintSetupState as BlueprintSetupState  # noqa
from .sansio.scaffold import _sentinel

if t.TYPE_CHECKING:  # pragma: no cover
    from .wrappers import Response


class Blueprint(SansioBlueprint):
    def __init__(
        self,
        name: str,
        import_name: str,
        static_folder: str | os.PathLike[str] | None = None,
        static_url_path: str | None = None,
        template_folder: str | os.PathLike[str] | None = None,
        url_prefix: str | None = None,
        subdomain: str | None = None,
        url_defaults: dict[str, t.Any] | None = None,
        root_path: str | None = None,
        cli_group: str | None = _sentinel,  # type: ignore
    ) -> None:
        super().__init__(
            name,
            import_name,
            static_folder,
            static_url_path,
            template_folder,
            url_prefix,
            subdomain,
            url_defaults,
            root_path,
            cli_group,
        )

        #: The Click command group for registering CLI commands for this
        #: object. The commands are available from the ``flask`` command
        #: once the application has been discovered and blueprints have
        #: been registered.
        self.cli = AppGroup()

        # Set the name of the Click group in case someone wants to add
        # the app's commands to another CLI tool.
        self.cli.name = self.name

    def get_send_file_max_age(self, filename: str | None) -> int | None:
        """Used by :func:`send_file` to determine the ``max_age`` cache
        value for a given file path if it wasn't passed.

        By default, this returns :data:`SEND_FILE_MAX_AGE_DEFAULT` from
        the configuration of :data:`~flask.current_app`. This defaults
        to ``None``, which tells the browser to use conditional requests
        instead of a timed cache, which is usually preferable.

        Note this is a duplicate of the same method in the Flask
        class.

        .. versionchanged:: 2.0
            The default configuration is ``None`` instead of 12 hours.

        .. versionadded:: 0.9
        """
        value = current_app.config["SEND_FILE_MAX_AGE_DEFAULT"]

        if value is None:
            return None

        if isinstance(value, timedelta):
            return int(value.total_seconds())

        return value  # type: ignore[no-any-return]

    def send_static_file(self, filename: str) -> Response:
        """The view function used to serve files from
        :attr:`static_folder`. A route is automatically registered for
        this view at :attr:`static_url_path` if :attr:`static_folder` is
        set.

        Note this is a duplicate of the same method in the Flask
        class.

        .. versionadded:: 0.5

        """
        if not self.has_static_folder:
            raise RuntimeError("'static_folder' must be set to serve static_files.")

        # send_file only knows to call get_send_file_max_age on the app,
        # call it here so it works for blueprints too.
        max_age = self.get_send_file_max_age(filename)
        return send_from_directory(
            t.cast(str, self.static_folder), filename, max_age=max_age
        )

    def open_resource(
        self, resource: str, mode: str = "rb", encoding: str | None = "utf-8"
    ) -> t.IO[t.AnyStr]:
        """Open a resource file relative to :attr:`root_path` for reading. The
        blueprint-relative equivalent of the app's :meth:`~.Flask.open_resource`
        method.

        :param resource: Path to the resource relative to :attr:`root_path`.
        :param mode: Open the file in this mode. Only reading is supported,
            valid values are ``"r"`` (or ``"rt"``) and ``"rb"``.
        :param encoding: Open the file with this encoding when opening in text
            mode. This is ignored when opening in binary mode.

        .. versionchanged:: 3.1
            Added the ``encoding`` parameter.
        """
        if mode not in {"r", "rt", "rb"}:
            raise ValueError("Resources can only be opened for reading.")

        path = os.path.join(self.root_path, resource)

        if mode == "rb":
            return open(path, mode)  # pyright: ignore

        return open(path, mode, encoding=encoding)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\google\protobuf\internal\testing_refleaks.py ===
# Protocol Buffers - Google's data interchange format
# Copyright 2008 Google Inc.  All rights reserved.
#
# Use of this source code is governed by a BSD-style
# license that can be found in the LICENSE file or at
# https://developers.google.com/open-source/licenses/bsd

"""A subclass of unittest.TestCase which checks for reference leaks.

To use:
- Use testing_refleak.BaseTestCase instead of unittest.TestCase
- Configure and compile Python with --with-pydebug

If sys.gettotalrefcount() is not available (because Python was built without
the Py_DEBUG option), then this module is a no-op and tests will run normally.
"""

import copyreg
import gc
import sys
import unittest


class LocalTestResult(unittest.TestResult):
  """A TestResult which forwards events to a parent object, except for Skips."""

  def __init__(self, parent_result):
    unittest.TestResult.__init__(self)
    self.parent_result = parent_result

  def addError(self, test, error):
    self.parent_result.addError(test, error)

  def addFailure(self, test, error):
    self.parent_result.addFailure(test, error)

  def addSkip(self, test, reason):
    pass

  def addDuration(self, test, duration):
    pass


class ReferenceLeakCheckerMixin(object):
  """A mixin class for TestCase, which checks reference counts."""

  NB_RUNS = 3

  def run(self, result=None):
    testMethod = getattr(self, self._testMethodName)
    expecting_failure_method = getattr(testMethod, "__unittest_expecting_failure__", False)
    expecting_failure_class = getattr(self, "__unittest_expecting_failure__", False)
    if expecting_failure_class or expecting_failure_method:
      return

    # python_message.py registers all Message classes to some pickle global
    # registry, which makes the classes immortal.
    # We save a copy of this registry, and reset it before we could references.
    self._saved_pickle_registry = copyreg.dispatch_table.copy()

    # Run the test twice, to warm up the instance attributes.
    super(ReferenceLeakCheckerMixin, self).run(result=result)
    super(ReferenceLeakCheckerMixin, self).run(result=result)

    local_result = LocalTestResult(result)
    num_flakes = 0
    refcount_deltas = []

    # Observe the refcount, then create oldrefcount which actually makes the
    # refcount 1 higher than the recorded value immediately
    oldrefcount = self._getRefcounts()
    while len(refcount_deltas) < self.NB_RUNS:
      oldrefcount = self._getRefcounts()
      super(ReferenceLeakCheckerMixin, self).run(result=local_result)
      newrefcount = self._getRefcounts()
      # If the GC was able to collect some objects after the call to run() that
      # it could not collect before the call, then the counts won't match.
      if newrefcount < oldrefcount and num_flakes < 2:
        # This result is (probably) a flake -- garbage collectors aren't very
        # predictable, but a lower ending refcount is the opposite of the
        # failure we are testing for. If the result is repeatable, then we will
        # eventually report it, but not after trying to eliminate it.
        num_flakes += 1
        continue
      num_flakes = 0
      refcount_deltas.append(newrefcount - oldrefcount)
    print(refcount_deltas, self)

    try:
      self.assertEqual(refcount_deltas, [0] * self.NB_RUNS)
    except Exception:  # pylint: disable=broad-except
      result.addError(self, sys.exc_info())

  def _getRefcounts(self):
    if hasattr(sys, "_clear_internal_caches"):  # Since 3.13
      sys._clear_internal_caches()  # pylint: disable=protected-access
    else:
      sys._clear_type_cache()  # pylint: disable=protected-access
    copyreg.dispatch_table.clear()
    copyreg.dispatch_table.update(self._saved_pickle_registry)
    # It is sometimes necessary to gc.collect() multiple times, to ensure
    # that all objects can be collected.
    gc.collect()
    gc.collect()
    gc.collect()
    return sys.gettotalrefcount()


if hasattr(sys, 'gettotalrefcount'):

  def TestCase(test_class):
    new_bases = (ReferenceLeakCheckerMixin,) + test_class.__bases__
    new_class = type(test_class)(
        test_class.__name__, new_bases, dict(test_class.__dict__))
    return new_class
  SkipReferenceLeakChecker = unittest.skip

else:
  # When PyDEBUG is not enabled, run the tests normally.

  def TestCase(test_class):
    return test_class

  def SkipReferenceLeakChecker(reason):
    del reason  # Don't skip, so don't need a reason.
    def Same(func):
      return func
    return Same

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\requests\status_codes.py ===
r"""
The ``codes`` object defines a mapping from common names for HTTP statuses
to their numerical codes, accessible either as attributes or as dictionary
items.

Example::

    >>> import requests
    >>> requests.codes['temporary_redirect']
    307
    >>> requests.codes.teapot
    418
    >>> requests.codes['\o/']
    200

Some codes have multiple names, and both upper- and lower-case versions of
the names are allowed. For example, ``codes.ok``, ``codes.OK``, and
``codes.okay`` all correspond to the HTTP status code 200.
"""

from .structures import LookupDict

_codes = {
    # Informational.
    100: ("continue",),
    101: ("switching_protocols",),
    102: ("processing", "early-hints"),
    103: ("checkpoint",),
    122: ("uri_too_long", "request_uri_too_long"),
    200: ("ok", "okay", "all_ok", "all_okay", "all_good", "\\o/", "✓"),
    201: ("created",),
    202: ("accepted",),
    203: ("non_authoritative_info", "non_authoritative_information"),
    204: ("no_content",),
    205: ("reset_content", "reset"),
    206: ("partial_content", "partial"),
    207: ("multi_status", "multiple_status", "multi_stati", "multiple_stati"),
    208: ("already_reported",),
    226: ("im_used",),
    # Redirection.
    300: ("multiple_choices",),
    301: ("moved_permanently", "moved", "\\o-"),
    302: ("found",),
    303: ("see_other", "other"),
    304: ("not_modified",),
    305: ("use_proxy",),
    306: ("switch_proxy",),
    307: ("temporary_redirect", "temporary_moved", "temporary"),
    308: (
        "permanent_redirect",
        "resume_incomplete",
        "resume",
    ),  # "resume" and "resume_incomplete" to be removed in 3.0
    # Client Error.
    400: ("bad_request", "bad"),
    401: ("unauthorized",),
    402: ("payment_required", "payment"),
    403: ("forbidden",),
    404: ("not_found", "-o-"),
    405: ("method_not_allowed", "not_allowed"),
    406: ("not_acceptable",),
    407: ("proxy_authentication_required", "proxy_auth", "proxy_authentication"),
    408: ("request_timeout", "timeout"),
    409: ("conflict",),
    410: ("gone",),
    411: ("length_required",),
    412: ("precondition_failed", "precondition"),
    413: ("request_entity_too_large", "content_too_large"),
    414: ("request_uri_too_large", "uri_too_long"),
    415: ("unsupported_media_type", "unsupported_media", "media_type"),
    416: (
        "requested_range_not_satisfiable",
        "requested_range",
        "range_not_satisfiable",
    ),
    417: ("expectation_failed",),
    418: ("im_a_teapot", "teapot", "i_am_a_teapot"),
    421: ("misdirected_request",),
    422: ("unprocessable_entity", "unprocessable", "unprocessable_content"),
    423: ("locked",),
    424: ("failed_dependency", "dependency"),
    425: ("unordered_collection", "unordered", "too_early"),
    426: ("upgrade_required", "upgrade"),
    428: ("precondition_required", "precondition"),
    429: ("too_many_requests", "too_many"),
    431: ("header_fields_too_large", "fields_too_large"),
    444: ("no_response", "none"),
    449: ("retry_with", "retry"),
    450: ("blocked_by_windows_parental_controls", "parental_controls"),
    451: ("unavailable_for_legal_reasons", "legal_reasons"),
    499: ("client_closed_request",),
    # Server Error.
    500: ("internal_server_error", "server_error", "/o\\", "✗"),
    501: ("not_implemented",),
    502: ("bad_gateway",),
    503: ("service_unavailable", "unavailable"),
    504: ("gateway_timeout",),
    505: ("http_version_not_supported", "http_version"),
    506: ("variant_also_negotiates",),
    507: ("insufficient_storage",),
    509: ("bandwidth_limit_exceeded", "bandwidth"),
    510: ("not_extended",),
    511: ("network_authentication_required", "network_auth", "network_authentication"),
}

codes = LookupDict(name="status_codes")


def _init():
    for code, titles in _codes.items():
        for title in titles:
            setattr(codes, title, code)
            if not title.startswith(("\\", "/")):
                setattr(codes, title.upper(), code)

    def doc(code):
        names = ", ".join(f"``{n}``" for n in _codes[code])
        return "* %d: %s" % (code, names)

    global __doc__
    __doc__ = (
        __doc__ + "\n" + "\n".join(doc(code) for code in sorted(_codes))
        if __doc__ is not None
        else None
    )


_init()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\requests\status_codes.py ===
r"""
The ``codes`` object defines a mapping from common names for HTTP statuses
to their numerical codes, accessible either as attributes or as dictionary
items.

Example::

    >>> import requests
    >>> requests.codes['temporary_redirect']
    307
    >>> requests.codes.teapot
    418
    >>> requests.codes['\o/']
    200

Some codes have multiple names, and both upper- and lower-case versions of
the names are allowed. For example, ``codes.ok``, ``codes.OK``, and
``codes.okay`` all correspond to the HTTP status code 200.
"""

from .structures import LookupDict

_codes = {
    # Informational.
    100: ("continue",),
    101: ("switching_protocols",),
    102: ("processing", "early-hints"),
    103: ("checkpoint",),
    122: ("uri_too_long", "request_uri_too_long"),
    200: ("ok", "okay", "all_ok", "all_okay", "all_good", "\\o/", "✓"),
    201: ("created",),
    202: ("accepted",),
    203: ("non_authoritative_info", "non_authoritative_information"),
    204: ("no_content",),
    205: ("reset_content", "reset"),
    206: ("partial_content", "partial"),
    207: ("multi_status", "multiple_status", "multi_stati", "multiple_stati"),
    208: ("already_reported",),
    226: ("im_used",),
    # Redirection.
    300: ("multiple_choices",),
    301: ("moved_permanently", "moved", "\\o-"),
    302: ("found",),
    303: ("see_other", "other"),
    304: ("not_modified",),
    305: ("use_proxy",),
    306: ("switch_proxy",),
    307: ("temporary_redirect", "temporary_moved", "temporary"),
    308: (
        "permanent_redirect",
        "resume_incomplete",
        "resume",
    ),  # "resume" and "resume_incomplete" to be removed in 3.0
    # Client Error.
    400: ("bad_request", "bad"),
    401: ("unauthorized",),
    402: ("payment_required", "payment"),
    403: ("forbidden",),
    404: ("not_found", "-o-"),
    405: ("method_not_allowed", "not_allowed"),
    406: ("not_acceptable",),
    407: ("proxy_authentication_required", "proxy_auth", "proxy_authentication"),
    408: ("request_timeout", "timeout"),
    409: ("conflict",),
    410: ("gone",),
    411: ("length_required",),
    412: ("precondition_failed", "precondition"),
    413: ("request_entity_too_large", "content_too_large"),
    414: ("request_uri_too_large", "uri_too_long"),
    415: ("unsupported_media_type", "unsupported_media", "media_type"),
    416: (
        "requested_range_not_satisfiable",
        "requested_range",
        "range_not_satisfiable",
    ),
    417: ("expectation_failed",),
    418: ("im_a_teapot", "teapot", "i_am_a_teapot"),
    421: ("misdirected_request",),
    422: ("unprocessable_entity", "unprocessable", "unprocessable_content"),
    423: ("locked",),
    424: ("failed_dependency", "dependency"),
    425: ("unordered_collection", "unordered", "too_early"),
    426: ("upgrade_required", "upgrade"),
    428: ("precondition_required", "precondition"),
    429: ("too_many_requests", "too_many"),
    431: ("header_fields_too_large", "fields_too_large"),
    444: ("no_response", "none"),
    449: ("retry_with", "retry"),
    450: ("blocked_by_windows_parental_controls", "parental_controls"),
    451: ("unavailable_for_legal_reasons", "legal_reasons"),
    499: ("client_closed_request",),
    # Server Error.
    500: ("internal_server_error", "server_error", "/o\\", "✗"),
    501: ("not_implemented",),
    502: ("bad_gateway",),
    503: ("service_unavailable", "unavailable"),
    504: ("gateway_timeout",),
    505: ("http_version_not_supported", "http_version"),
    506: ("variant_also_negotiates",),
    507: ("insufficient_storage",),
    509: ("bandwidth_limit_exceeded", "bandwidth"),
    510: ("not_extended",),
    511: ("network_authentication_required", "network_auth", "network_authentication"),
}

codes = LookupDict(name="status_codes")


def _init():
    for code, titles in _codes.items():
        for title in titles:
            setattr(codes, title, code)
            if not title.startswith(("\\", "/")):
                setattr(codes, title.upper(), code)

    def doc(code):
        names = ", ".join(f"``{n}``" for n in _codes[code])
        return "* %d: %s" % (code, names)

    global __doc__
    __doc__ = (
        __doc__ + "\n" + "\n".join(doc(code) for code in sorted(_codes))
        if __doc__ is not None
        else None
    )


_init()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\engine\_py_row.py ===
# engine/_py_row.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
from __future__ import annotations

import operator
import typing
from typing import Any
from typing import Callable
from typing import Dict
from typing import Iterator
from typing import List
from typing import Mapping
from typing import Optional
from typing import Tuple
from typing import Type

if typing.TYPE_CHECKING:
    from .result import _KeyType
    from .result import _ProcessorsType
    from .result import _RawRowType
    from .result import _TupleGetterType
    from .result import ResultMetaData

MD_INDEX = 0  # integer index in cursor.description


class BaseRow:
    __slots__ = ("_parent", "_data", "_key_to_index")

    _parent: ResultMetaData
    _key_to_index: Mapping[_KeyType, int]
    _data: _RawRowType

    def __init__(
        self,
        parent: ResultMetaData,
        processors: Optional[_ProcessorsType],
        key_to_index: Mapping[_KeyType, int],
        data: _RawRowType,
    ):
        """Row objects are constructed by CursorResult objects."""
        object.__setattr__(self, "_parent", parent)

        object.__setattr__(self, "_key_to_index", key_to_index)

        if processors:
            object.__setattr__(
                self,
                "_data",
                tuple(
                    [
                        proc(value) if proc else value
                        for proc, value in zip(processors, data)
                    ]
                ),
            )
        else:
            object.__setattr__(self, "_data", tuple(data))

    def __reduce__(self) -> Tuple[Callable[..., BaseRow], Tuple[Any, ...]]:
        return (
            rowproxy_reconstructor,
            (self.__class__, self.__getstate__()),
        )

    def __getstate__(self) -> Dict[str, Any]:
        return {"_parent": self._parent, "_data": self._data}

    def __setstate__(self, state: Dict[str, Any]) -> None:
        parent = state["_parent"]
        object.__setattr__(self, "_parent", parent)
        object.__setattr__(self, "_data", state["_data"])
        object.__setattr__(self, "_key_to_index", parent._key_to_index)

    def _values_impl(self) -> List[Any]:
        return list(self)

    def __iter__(self) -> Iterator[Any]:
        return iter(self._data)

    def __len__(self) -> int:
        return len(self._data)

    def __hash__(self) -> int:
        return hash(self._data)

    def __getitem__(self, key: Any) -> Any:
        return self._data[key]

    def _get_by_key_impl_mapping(self, key: str) -> Any:
        try:
            return self._data[self._key_to_index[key]]
        except KeyError:
            pass
        self._parent._key_not_found(key, False)

    def __getattr__(self, name: str) -> Any:
        try:
            return self._data[self._key_to_index[name]]
        except KeyError:
            pass
        self._parent._key_not_found(name, True)

    def _to_tuple_instance(self) -> Tuple[Any, ...]:
        return self._data


# This reconstructor is necessary so that pickles with the Cy extension or
# without use the same Binary format.
def rowproxy_reconstructor(
    cls: Type[BaseRow], state: Dict[str, Any]
) -> BaseRow:
    obj = cls.__new__(cls)
    obj.__setstate__(state)
    return obj


def tuplegetter(*indexes: int) -> _TupleGetterType:
    if len(indexes) != 1:
        for i in range(1, len(indexes)):
            if indexes[i - 1] != indexes[i] - 1:
                return operator.itemgetter(*indexes)
    # slice form is faster but returns a list if input is list
    return operator.itemgetter(slice(indexes[0], indexes[-1] + 1))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\flask_wtf\form.py ===
from flask import current_app
from flask import request
from flask import session
from markupsafe import Markup
from werkzeug.datastructures import CombinedMultiDict
from werkzeug.datastructures import ImmutableMultiDict
from werkzeug.utils import cached_property
from wtforms import Form
from wtforms.meta import DefaultMeta
from wtforms.widgets import HiddenInput

from .csrf import _FlaskFormCSRF

try:
    from .i18n import translations
except ImportError:
    translations = None  # babel not installed


SUBMIT_METHODS = {"POST", "PUT", "PATCH", "DELETE"}
_Auto = object()


class FlaskForm(Form):
    """Flask-specific subclass of WTForms :class:`~wtforms.form.Form`.

    If ``formdata`` is not specified, this will use :attr:`flask.request.form`
    and :attr:`flask.request.files`.  Explicitly pass ``formdata=None`` to
    prevent this.
    """

    class Meta(DefaultMeta):
        csrf_class = _FlaskFormCSRF
        csrf_context = session  # not used, provided for custom csrf_class

        @cached_property
        def csrf(self):
            return current_app.config.get("WTF_CSRF_ENABLED", True)

        @cached_property
        def csrf_secret(self):
            return current_app.config.get("WTF_CSRF_SECRET_KEY", current_app.secret_key)

        @cached_property
        def csrf_field_name(self):
            return current_app.config.get("WTF_CSRF_FIELD_NAME", "csrf_token")

        @cached_property
        def csrf_time_limit(self):
            return current_app.config.get("WTF_CSRF_TIME_LIMIT", 3600)

        def wrap_formdata(self, form, formdata):
            if formdata is _Auto:
                if _is_submitted():
                    if request.files:
                        return CombinedMultiDict((request.files, request.form))
                    elif request.form:
                        return request.form
                    elif request.is_json:
                        return ImmutableMultiDict(request.get_json())

                return None

            return formdata

        def get_translations(self, form):
            if not current_app.config.get("WTF_I18N_ENABLED", True):
                return super().get_translations(form)

            return translations

    def __init__(self, formdata=_Auto, **kwargs):
        super().__init__(formdata=formdata, **kwargs)

    def is_submitted(self):
        """Consider the form submitted if there is an active request and
        the method is ``POST``, ``PUT``, ``PATCH``, or ``DELETE``.
        """

        return _is_submitted()

    def validate_on_submit(self, extra_validators=None):
        """Call :meth:`validate` only if the form is submitted.
        This is a shortcut for ``form.is_submitted() and form.validate()``.
        """
        return self.is_submitted() and self.validate(extra_validators=extra_validators)

    def hidden_tag(self, *fields):
        """Render the form's hidden fields in one call.

        A field is considered hidden if it uses the
        :class:`~wtforms.widgets.HiddenInput` widget.

        If ``fields`` are given, only render the given fields that
        are hidden.  If a string is passed, render the field with that
        name if it exists.

        .. versionchanged:: 0.13

           No longer wraps inputs in hidden div.
           This is valid HTML 5.

        .. versionchanged:: 0.13

           Skip passed fields that aren't hidden.
           Skip passed names that don't exist.
        """

        def hidden_fields(fields):
            for f in fields:
                if isinstance(f, str):
                    f = getattr(self, f, None)

                if f is None or not isinstance(f.widget, HiddenInput):
                    continue

                yield f

        return Markup("\n".join(str(f) for f in hidden_fields(fields or self)))


def _is_submitted():
    """Consider the form submitted if there is an active request and
    the method is ``POST``, ``PUT``, ``PATCH``, or ``DELETE``.
    """

    return bool(request) and request.method in SUBMIT_METHODS

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\chart\label.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Sequence,
    Alias,
    Typed
)
from openpyxl.descriptors.excel import ExtensionList
from openpyxl.descriptors.nested import (
    NestedNoneSet,
    NestedBool,
    NestedString,
    NestedInteger,
    )

from .shapes import GraphicalProperties
from .text import RichText


class _DataLabelBase(Serialisable):

    numFmt = NestedString(allow_none=True, attribute="formatCode")
    spPr = Typed(expected_type=GraphicalProperties, allow_none=True)
    graphicalProperties = Alias('spPr')
    txPr = Typed(expected_type=RichText, allow_none=True)
    textProperties = Alias('txPr')
    dLblPos = NestedNoneSet(values=['bestFit', 'b', 'ctr', 'inBase', 'inEnd',
                                    'l', 'outEnd', 'r', 't'])
    position = Alias('dLblPos')
    showLegendKey = NestedBool(allow_none=True)
    showVal = NestedBool(allow_none=True)
    showCatName = NestedBool(allow_none=True)
    showSerName = NestedBool(allow_none=True)
    showPercent = NestedBool(allow_none=True)
    showBubbleSize = NestedBool(allow_none=True)
    showLeaderLines = NestedBool(allow_none=True)
    separator = NestedString(allow_none=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ("numFmt", "spPr", "txPr", "dLblPos", "showLegendKey",
                    "showVal", "showCatName", "showSerName", "showPercent", "showBubbleSize",
                    "showLeaderLines", "separator")

    def __init__(self,
                 numFmt=None,
                 spPr=None,
                 txPr=None,
                 dLblPos=None,
                 showLegendKey=None,
                 showVal=None,
                 showCatName=None,
                 showSerName=None,
                 showPercent=None,
                 showBubbleSize=None,
                 showLeaderLines=None,
                 separator=None,
                 extLst=None,
                 ):
        self.numFmt = numFmt
        self.spPr = spPr
        self.txPr = txPr
        self.dLblPos = dLblPos
        self.showLegendKey = showLegendKey
        self.showVal = showVal
        self.showCatName = showCatName
        self.showSerName = showSerName
        self.showPercent = showPercent
        self.showBubbleSize = showBubbleSize
        self.showLeaderLines = showLeaderLines
        self.separator = separator


class DataLabel(_DataLabelBase):

    tagname = "dLbl"

    idx = NestedInteger()

    numFmt = _DataLabelBase.numFmt
    spPr = _DataLabelBase.spPr
    txPr = _DataLabelBase.txPr
    dLblPos = _DataLabelBase.dLblPos
    showLegendKey = _DataLabelBase.showLegendKey
    showVal = _DataLabelBase.showVal
    showCatName = _DataLabelBase.showCatName
    showSerName = _DataLabelBase.showSerName
    showPercent = _DataLabelBase.showPercent
    showBubbleSize = _DataLabelBase.showBubbleSize
    showLeaderLines = _DataLabelBase.showLeaderLines
    separator = _DataLabelBase.separator
    extLst = _DataLabelBase.extLst

    __elements__ = ("idx",)  + _DataLabelBase.__elements__

    def __init__(self, idx=0, **kw ):
        self.idx = idx
        super().__init__(**kw)


class DataLabelList(_DataLabelBase):

    tagname = "dLbls"

    dLbl = Sequence(expected_type=DataLabel, allow_none=True)

    delete = NestedBool(allow_none=True)
    numFmt = _DataLabelBase.numFmt
    spPr = _DataLabelBase.spPr
    txPr = _DataLabelBase.txPr
    dLblPos = _DataLabelBase.dLblPos
    showLegendKey = _DataLabelBase.showLegendKey
    showVal = _DataLabelBase.showVal
    showCatName = _DataLabelBase.showCatName
    showSerName = _DataLabelBase.showSerName
    showPercent = _DataLabelBase.showPercent
    showBubbleSize = _DataLabelBase.showBubbleSize
    showLeaderLines = _DataLabelBase.showLeaderLines
    separator = _DataLabelBase.separator
    extLst = _DataLabelBase.extLst

    __elements__ = ("delete", "dLbl",) + _DataLabelBase.__elements__

    def __init__(self, dLbl=(), delete=None,  **kw):
        self.dLbl = dLbl
        self.delete = delete
        super().__init__(**kw)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\io\excel\_pyxlsb.py ===
# pyright: reportMissingImports=false
from __future__ import annotations

from typing import TYPE_CHECKING

from pandas.compat._optional import import_optional_dependency
from pandas.util._decorators import doc

from pandas.core.shared_docs import _shared_docs

from pandas.io.excel._base import BaseExcelReader

if TYPE_CHECKING:
    from pyxlsb import Workbook

    from pandas._typing import (
        FilePath,
        ReadBuffer,
        Scalar,
        StorageOptions,
    )


class PyxlsbReader(BaseExcelReader["Workbook"]):
    @doc(storage_options=_shared_docs["storage_options"])
    def __init__(
        self,
        filepath_or_buffer: FilePath | ReadBuffer[bytes],
        storage_options: StorageOptions | None = None,
        engine_kwargs: dict | None = None,
    ) -> None:
        """
        Reader using pyxlsb engine.

        Parameters
        ----------
        filepath_or_buffer : str, path object, or Workbook
            Object to be parsed.
        {storage_options}
        engine_kwargs : dict, optional
            Arbitrary keyword arguments passed to excel engine.
        """
        import_optional_dependency("pyxlsb")
        # This will call load_workbook on the filepath or buffer
        # And set the result to the book-attribute
        super().__init__(
            filepath_or_buffer,
            storage_options=storage_options,
            engine_kwargs=engine_kwargs,
        )

    @property
    def _workbook_class(self) -> type[Workbook]:
        from pyxlsb import Workbook

        return Workbook

    def load_workbook(
        self, filepath_or_buffer: FilePath | ReadBuffer[bytes], engine_kwargs
    ) -> Workbook:
        from pyxlsb import open_workbook

        # TODO: hack in buffer capability
        # This might need some modifications to the Pyxlsb library
        # Actual work for opening it is in xlsbpackage.py, line 20-ish

        return open_workbook(filepath_or_buffer, **engine_kwargs)

    @property
    def sheet_names(self) -> list[str]:
        return self.book.sheets

    def get_sheet_by_name(self, name: str):
        self.raise_if_bad_sheet_by_name(name)
        return self.book.get_sheet(name)

    def get_sheet_by_index(self, index: int):
        self.raise_if_bad_sheet_by_index(index)
        # pyxlsb sheets are indexed from 1 onwards
        # There's a fix for this in the source, but the pypi package doesn't have it
        return self.book.get_sheet(index + 1)

    def _convert_cell(self, cell) -> Scalar:
        # TODO: there is no way to distinguish between floats and datetimes in pyxlsb
        # This means that there is no way to read datetime types from an xlsb file yet
        if cell.v is None:
            return ""  # Prevents non-named columns from not showing up as Unnamed: i
        if isinstance(cell.v, float):
            val = int(cell.v)
            if val == cell.v:
                return val
            else:
                return float(cell.v)

        return cell.v

    def get_sheet_data(
        self,
        sheet,
        file_rows_needed: int | None = None,
    ) -> list[list[Scalar]]:
        data: list[list[Scalar]] = []
        previous_row_number = -1
        # When sparse=True the rows can have different lengths and empty rows are
        # not returned. The cells are namedtuples of row, col, value (r, c, v).
        for row in sheet.rows(sparse=True):
            row_number = row[0].r
            converted_row = [self._convert_cell(cell) for cell in row]
            while converted_row and converted_row[-1] == "":
                # trim trailing empty elements
                converted_row.pop()
            if converted_row:
                data.extend([[]] * (row_number - previous_row_number - 1))
                data.append(converted_row)
                previous_row_number = row_number
            if file_rows_needed is not None and len(data) >= file_rows_needed:
                break
        if data:
            # extend rows to max_width
            max_width = max(len(data_row) for data_row in data)
            if min(len(data_row) for data_row in data) < max_width:
                empty_cell: list[Scalar] = [""]
                data = [
                    data_row + (max_width - len(data_row)) * empty_cell
                    for data_row in data
                ]
        return data

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\models\search_scope.py ===
import itertools
import logging
import os
import posixpath
import urllib.parse
from dataclasses import dataclass
from typing import List

from pip._vendor.packaging.utils import canonicalize_name

from pip._internal.models.index import PyPI
from pip._internal.utils.compat import has_tls
from pip._internal.utils.misc import normalize_path, redact_auth_from_url

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class SearchScope:
    """
    Encapsulates the locations that pip is configured to search.
    """

    __slots__ = ["find_links", "index_urls", "no_index"]

    find_links: List[str]
    index_urls: List[str]
    no_index: bool

    @classmethod
    def create(
        cls,
        find_links: List[str],
        index_urls: List[str],
        no_index: bool,
    ) -> "SearchScope":
        """
        Create a SearchScope object after normalizing the `find_links`.
        """
        # Build find_links. If an argument starts with ~, it may be
        # a local file relative to a home directory. So try normalizing
        # it and if it exists, use the normalized version.
        # This is deliberately conservative - it might be fine just to
        # blindly normalize anything starting with a ~...
        built_find_links: List[str] = []
        for link in find_links:
            if link.startswith("~"):
                new_link = normalize_path(link)
                if os.path.exists(new_link):
                    link = new_link
            built_find_links.append(link)

        # If we don't have TLS enabled, then WARN if anyplace we're looking
        # relies on TLS.
        if not has_tls():
            for link in itertools.chain(index_urls, built_find_links):
                parsed = urllib.parse.urlparse(link)
                if parsed.scheme == "https":
                    logger.warning(
                        "pip is configured with locations that require "
                        "TLS/SSL, however the ssl module in Python is not "
                        "available."
                    )
                    break

        return cls(
            find_links=built_find_links,
            index_urls=index_urls,
            no_index=no_index,
        )

    def get_formatted_locations(self) -> str:
        lines = []
        redacted_index_urls = []
        if self.index_urls and self.index_urls != [PyPI.simple_url]:
            for url in self.index_urls:
                redacted_index_url = redact_auth_from_url(url)

                # Parse the URL
                purl = urllib.parse.urlsplit(redacted_index_url)

                # URL is generally invalid if scheme and netloc is missing
                # there are issues with Python and URL parsing, so this test
                # is a bit crude. See bpo-20271, bpo-23505. Python doesn't
                # always parse invalid URLs correctly - it should raise
                # exceptions for malformed URLs
                if not purl.scheme and not purl.netloc:
                    logger.warning(
                        'The index url "%s" seems invalid, please provide a scheme.',
                        redacted_index_url,
                    )

                redacted_index_urls.append(redacted_index_url)

            lines.append(
                "Looking in indexes: {}".format(", ".join(redacted_index_urls))
            )

        if self.find_links:
            lines.append(
                "Looking in links: {}".format(
                    ", ".join(redact_auth_from_url(url) for url in self.find_links)
                )
            )
        return "\n".join(lines)

    def get_index_urls_locations(self, project_name: str) -> List[str]:
        """Returns the locations found via self.index_urls

        Checks the url_name on the main (first in the list) index and
        use this url_name to produce all locations
        """

        def mkurl_pypi_url(url: str) -> str:
            loc = posixpath.join(
                url, urllib.parse.quote(canonicalize_name(project_name))
            )
            # For maximum compatibility with easy_install, ensure the path
            # ends in a trailing slash.  Although this isn't in the spec
            # (and PyPI can handle it without the slash) some other index
            # implementations might break if they relied on easy_install's
            # behavior.
            if not loc.endswith("/"):
                loc = loc + "/"
            return loc

        return [mkurl_pypi_url(url) for url in self.index_urls]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\requests\help.py ===
"""Module containing bug report helper(s)."""

import json
import platform
import ssl
import sys

from pip._vendor import idna
from pip._vendor import urllib3

from . import __version__ as requests_version

charset_normalizer = None
chardet = None

try:
    from pip._vendor.urllib3.contrib import pyopenssl
except ImportError:
    pyopenssl = None
    OpenSSL = None
    cryptography = None
else:
    import cryptography
    import OpenSSL


def _implementation():
    """Return a dict with the Python implementation and version.

    Provide both the name and the version of the Python implementation
    currently running. For example, on CPython 3.10.3 it will return
    {'name': 'CPython', 'version': '3.10.3'}.

    This function works best on CPython and PyPy: in particular, it probably
    doesn't work for Jython or IronPython. Future investigation should be done
    to work out the correct shape of the code for those platforms.
    """
    implementation = platform.python_implementation()

    if implementation == "CPython":
        implementation_version = platform.python_version()
    elif implementation == "PyPy":
        implementation_version = "{}.{}.{}".format(
            sys.pypy_version_info.major,
            sys.pypy_version_info.minor,
            sys.pypy_version_info.micro,
        )
        if sys.pypy_version_info.releaselevel != "final":
            implementation_version = "".join(
                [implementation_version, sys.pypy_version_info.releaselevel]
            )
    elif implementation == "Jython":
        implementation_version = platform.python_version()  # Complete Guess
    elif implementation == "IronPython":
        implementation_version = platform.python_version()  # Complete Guess
    else:
        implementation_version = "Unknown"

    return {"name": implementation, "version": implementation_version}


def info():
    """Generate information for a bug report."""
    try:
        platform_info = {
            "system": platform.system(),
            "release": platform.release(),
        }
    except OSError:
        platform_info = {
            "system": "Unknown",
            "release": "Unknown",
        }

    implementation_info = _implementation()
    urllib3_info = {"version": urllib3.__version__}
    charset_normalizer_info = {"version": None}
    chardet_info = {"version": None}
    if charset_normalizer:
        charset_normalizer_info = {"version": charset_normalizer.__version__}
    if chardet:
        chardet_info = {"version": chardet.__version__}

    pyopenssl_info = {
        "version": None,
        "openssl_version": "",
    }
    if OpenSSL:
        pyopenssl_info = {
            "version": OpenSSL.__version__,
            "openssl_version": f"{OpenSSL.SSL.OPENSSL_VERSION_NUMBER:x}",
        }
    cryptography_info = {
        "version": getattr(cryptography, "__version__", ""),
    }
    idna_info = {
        "version": getattr(idna, "__version__", ""),
    }

    system_ssl = ssl.OPENSSL_VERSION_NUMBER
    system_ssl_info = {"version": f"{system_ssl:x}" if system_ssl is not None else ""}

    return {
        "platform": platform_info,
        "implementation": implementation_info,
        "system_ssl": system_ssl_info,
        "using_pyopenssl": pyopenssl is not None,
        "using_charset_normalizer": chardet is None,
        "pyOpenSSL": pyopenssl_info,
        "urllib3": urllib3_info,
        "chardet": chardet_info,
        "charset_normalizer": charset_normalizer_info,
        "cryptography": cryptography_info,
        "idna": idna_info,
        "requests": {
            "version": requests_version,
        },
    }


def main():
    """Pretty-print the bug information as JSON."""
    print(json.dumps(info(), sort_keys=True, indent=2))


if __name__ == "__main__":
    main()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\tools\ort_format_model\ort_flatbuffers_py\fbs\NodeEdge.py ===
# automatically generated by the FlatBuffers compiler, do not modify

# namespace: fbs

import flatbuffers
from flatbuffers.compat import import_numpy
np = import_numpy()

class NodeEdge(object):
    __slots__ = ['_tab']

    @classmethod
    def GetRootAs(cls, buf, offset=0):
        n = flatbuffers.encode.Get(flatbuffers.packer.uoffset, buf, offset)
        x = NodeEdge()
        x.Init(buf, n + offset)
        return x

    @classmethod
    def GetRootAsNodeEdge(cls, buf, offset=0):
        """This method is deprecated. Please switch to GetRootAs."""
        return cls.GetRootAs(buf, offset)
    @classmethod
    def NodeEdgeBufferHasIdentifier(cls, buf, offset, size_prefixed=False):
        return flatbuffers.util.BufferHasIdentifier(buf, offset, b"\x4F\x52\x54\x4D", size_prefixed=size_prefixed)

    # NodeEdge
    def Init(self, buf, pos):
        self._tab = flatbuffers.table.Table(buf, pos)

    # NodeEdge
    def NodeIndex(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(4))
        if o != 0:
            return self._tab.Get(flatbuffers.number_types.Uint32Flags, o + self._tab.Pos)
        return 0

    # NodeEdge
    def InputEdges(self, j):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        if o != 0:
            x = self._tab.Vector(o)
            x += flatbuffers.number_types.UOffsetTFlags.py_type(j) * 12
            from ort_flatbuffers_py.fbs.EdgeEnd import EdgeEnd
            obj = EdgeEnd()
            obj.Init(self._tab.Bytes, x)
            return obj
        return None

    # NodeEdge
    def InputEdgesLength(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        if o != 0:
            return self._tab.VectorLen(o)
        return 0

    # NodeEdge
    def InputEdgesIsNone(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        return o == 0

    # NodeEdge
    def OutputEdges(self, j):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(8))
        if o != 0:
            x = self._tab.Vector(o)
            x += flatbuffers.number_types.UOffsetTFlags.py_type(j) * 12
            from ort_flatbuffers_py.fbs.EdgeEnd import EdgeEnd
            obj = EdgeEnd()
            obj.Init(self._tab.Bytes, x)
            return obj
        return None

    # NodeEdge
    def OutputEdgesLength(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(8))
        if o != 0:
            return self._tab.VectorLen(o)
        return 0

    # NodeEdge
    def OutputEdgesIsNone(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(8))
        return o == 0

def NodeEdgeStart(builder):
    builder.StartObject(3)

def Start(builder):
    NodeEdgeStart(builder)

def NodeEdgeAddNodeIndex(builder, nodeIndex):
    builder.PrependUint32Slot(0, nodeIndex, 0)

def AddNodeIndex(builder, nodeIndex):
    NodeEdgeAddNodeIndex(builder, nodeIndex)

def NodeEdgeAddInputEdges(builder, inputEdges):
    builder.PrependUOffsetTRelativeSlot(1, flatbuffers.number_types.UOffsetTFlags.py_type(inputEdges), 0)

def AddInputEdges(builder, inputEdges):
    NodeEdgeAddInputEdges(builder, inputEdges)

def NodeEdgeStartInputEdgesVector(builder, numElems):
    return builder.StartVector(12, numElems, 4)

def StartInputEdgesVector(builder, numElems: int) -> int:
    return NodeEdgeStartInputEdgesVector(builder, numElems)

def NodeEdgeAddOutputEdges(builder, outputEdges):
    builder.PrependUOffsetTRelativeSlot(2, flatbuffers.number_types.UOffsetTFlags.py_type(outputEdges), 0)

def AddOutputEdges(builder, outputEdges):
    NodeEdgeAddOutputEdges(builder, outputEdges)

def NodeEdgeStartOutputEdgesVector(builder, numElems):
    return builder.StartVector(12, numElems, 4)

def StartOutputEdgesVector(builder, numElems: int) -> int:
    return NodeEdgeStartOutputEdgesVector(builder, numElems)

def NodeEdgeEnd(builder):
    return builder.EndObject()

def End(builder):
    return NodeEdgeEnd(builder)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\mssql\pymssql.py ===
# dialects/mssql/pymssql.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors


"""
.. dialect:: mssql+pymssql
    :name: pymssql
    :dbapi: pymssql
    :connectstring: mssql+pymssql://<username>:<password>@<freetds_name>/?charset=utf8

pymssql is a Python module that provides a Python DBAPI interface around
`FreeTDS <https://www.freetds.org/>`_.

.. versionchanged:: 2.0.5

    pymssql was restored to SQLAlchemy's continuous integration testing


"""  # noqa
import re

from .base import MSDialect
from .base import MSIdentifierPreparer
from ... import types as sqltypes
from ... import util
from ...engine import processors


class _MSNumeric_pymssql(sqltypes.Numeric):
    def result_processor(self, dialect, type_):
        if not self.asdecimal:
            return processors.to_float
        else:
            return sqltypes.Numeric.result_processor(self, dialect, type_)


class MSIdentifierPreparer_pymssql(MSIdentifierPreparer):
    def __init__(self, dialect):
        super().__init__(dialect)
        # pymssql has the very unusual behavior that it uses pyformat
        # yet does not require that percent signs be doubled
        self._double_percents = False


class MSDialect_pymssql(MSDialect):
    supports_statement_cache = True
    supports_native_decimal = True
    supports_native_uuid = True
    driver = "pymssql"

    preparer = MSIdentifierPreparer_pymssql

    colspecs = util.update_copy(
        MSDialect.colspecs,
        {sqltypes.Numeric: _MSNumeric_pymssql, sqltypes.Float: sqltypes.Float},
    )

    @classmethod
    def import_dbapi(cls):
        module = __import__("pymssql")
        # pymmsql < 2.1.1 doesn't have a Binary method.  we use string
        client_ver = tuple(int(x) for x in module.__version__.split("."))
        if client_ver < (2, 1, 1):
            # TODO: monkeypatching here is less than ideal
            module.Binary = lambda x: x if hasattr(x, "decode") else str(x)

        if client_ver < (1,):
            util.warn(
                "The pymssql dialect expects at least "
                "the 1.0 series of the pymssql DBAPI."
            )
        return module

    def _get_server_version_info(self, connection):
        vers = connection.exec_driver_sql("select @@version").scalar()
        m = re.match(r"Microsoft .*? - (\d+)\.(\d+)\.(\d+)\.(\d+)", vers)
        if m:
            return tuple(int(x) for x in m.group(1, 2, 3, 4))
        else:
            return None

    def create_connect_args(self, url):
        opts = url.translate_connect_args(username="user")
        opts.update(url.query)
        port = opts.pop("port", None)
        if port and "host" in opts:
            opts["host"] = "%s:%s" % (opts["host"], port)
        return ([], opts)

    def is_disconnect(self, e, connection, cursor):
        for msg in (
            "Adaptive Server connection timed out",
            "Net-Lib error during Connection reset by peer",
            "message 20003",  # connection timeout
            "Error 10054",
            "Not connected to any MS SQL server",
            "Connection is closed",
            "message 20006",  # Write to the server failed
            "message 20017",  # Unexpected EOF from the server
            "message 20047",  # DBPROCESS is dead or not enabled
            "The server failed to resume the transaction",
        ):
            if msg in str(e):
                return True
        else:
            return False

    def get_isolation_level_values(self, dbapi_connection):
        return super().get_isolation_level_values(dbapi_connection) + [
            "AUTOCOMMIT"
        ]

    def set_isolation_level(self, dbapi_connection, level):
        if level == "AUTOCOMMIT":
            dbapi_connection.autocommit(True)
        else:
            dbapi_connection.autocommit(False)
            super().set_isolation_level(dbapi_connection, level)


dialect = MSDialect_pymssql

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\assumptions\predicates\ntheory.py ===
from sympy.assumptions import Predicate
from sympy.multipledispatch import Dispatcher


class PrimePredicate(Predicate):
    """
    Prime number predicate.

    Explanation
    ===========

    ``ask(Q.prime(x))`` is true iff ``x`` is a natural number greater
    than 1 that has no positive divisors other than ``1`` and the
    number itself.

    Examples
    ========

    >>> from sympy import Q, ask
    >>> ask(Q.prime(0))
    False
    >>> ask(Q.prime(1))
    False
    >>> ask(Q.prime(2))
    True
    >>> ask(Q.prime(20))
    False
    >>> ask(Q.prime(-3))
    False

    """
    name = 'prime'
    handler = Dispatcher(
        "PrimeHandler",
        doc=("Handler for key 'prime'. Test that an expression represents a prime"
        " number. When the expression is an exact number, the result (when True)"
        " is subject to the limitations of isprime() which is used to return the "
        "result.")
    )


class CompositePredicate(Predicate):
    """
    Composite number predicate.

    Explanation
    ===========

    ``ask(Q.composite(x))`` is true iff ``x`` is a positive integer and has
    at least one positive divisor other than ``1`` and the number itself.

    Examples
    ========

    >>> from sympy import Q, ask
    >>> ask(Q.composite(0))
    False
    >>> ask(Q.composite(1))
    False
    >>> ask(Q.composite(2))
    False
    >>> ask(Q.composite(20))
    True

    """
    name = 'composite'
    handler = Dispatcher("CompositeHandler", doc="Handler for key 'composite'.")


class EvenPredicate(Predicate):
    """
    Even number predicate.

    Explanation
    ===========

    ``ask(Q.even(x))`` is true iff ``x`` belongs to the set of even
    integers.

    Examples
    ========

    >>> from sympy import Q, ask, pi
    >>> ask(Q.even(0))
    True
    >>> ask(Q.even(2))
    True
    >>> ask(Q.even(3))
    False
    >>> ask(Q.even(pi))
    False

    """
    name = 'even'
    handler = Dispatcher("EvenHandler", doc="Handler for key 'even'.")


class OddPredicate(Predicate):
    """
    Odd number predicate.

    Explanation
    ===========

    ``ask(Q.odd(x))`` is true iff ``x`` belongs to the set of odd numbers.

    Examples
    ========

    >>> from sympy import Q, ask, pi
    >>> ask(Q.odd(0))
    False
    >>> ask(Q.odd(2))
    False
    >>> ask(Q.odd(3))
    True
    >>> ask(Q.odd(pi))
    False

    """
    name = 'odd'
    handler = Dispatcher(
        "OddHandler",
        doc=("Handler for key 'odd'. Test that an expression represents an odd"
        " number.")
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\solvers\decompogen.py ===
from sympy.core import (Function, Pow, sympify, Expr)
from sympy.core.relational import Relational
from sympy.core.singleton import S
from sympy.polys import Poly, decompose
from sympy.utilities.misc import func_name
from sympy.functions.elementary.miscellaneous import Min, Max


def decompogen(f, symbol):
    """
    Computes General functional decomposition of ``f``.
    Given an expression ``f``, returns a list ``[f_1, f_2, ..., f_n]``,
    where::
              f = f_1 o f_2 o ... f_n = f_1(f_2(... f_n))

    Note: This is a General decomposition function. It also decomposes
    Polynomials. For only Polynomial decomposition see ``decompose`` in polys.

    Examples
    ========

    >>> from sympy.abc import x
    >>> from sympy import decompogen, sqrt, sin, cos
    >>> decompogen(sin(cos(x)), x)
    [sin(x), cos(x)]
    >>> decompogen(sin(x)**2 + sin(x) + 1, x)
    [x**2 + x + 1, sin(x)]
    >>> decompogen(sqrt(6*x**2 - 5), x)
    [sqrt(x), 6*x**2 - 5]
    >>> decompogen(sin(sqrt(cos(x**2 + 1))), x)
    [sin(x), sqrt(x), cos(x), x**2 + 1]
    >>> decompogen(x**4 + 2*x**3 - x - 1, x)
    [x**2 - x - 1, x**2 + x]

    """
    f = sympify(f)
    if not isinstance(f, Expr) or isinstance(f, Relational):
        raise TypeError('expecting Expr but got: `%s`' % func_name(f))
    if symbol not in f.free_symbols:
        return [f]


    # ===== Simple Functions ===== #
    if isinstance(f, (Function, Pow)):
        if f.is_Pow and f.base == S.Exp1:
            arg = f.exp
        else:
            arg = f.args[0]
        if arg == symbol:
            return [f]
        return [f.subs(arg, symbol)] + decompogen(arg, symbol)

    # ===== Min/Max Functions ===== #
    if isinstance(f, (Min, Max)):
        args = list(f.args)
        d0 = None
        for i, a in enumerate(args):
            if not a.has_free(symbol):
                continue
            d = decompogen(a, symbol)
            if len(d) == 1:
                d = [symbol] + d
            if d0 is None:
                d0 = d[1:]
            elif d[1:] != d0:
                # decomposition is not the same for each arg:
                # mark as having no decomposition
                d = [symbol]
                break
            args[i] = d[0]
        if d[0] == symbol:
            return [f]
        return [f.func(*args)] + d0

    # ===== Convert to Polynomial ===== #
    fp = Poly(f)
    gens = list(filter(lambda x: symbol in x.free_symbols, fp.gens))

    if len(gens) == 1 and gens[0] != symbol:
        f1 = f.subs(gens[0], symbol)
        f2 = gens[0]
        return [f1] + decompogen(f2, symbol)

    # ===== Polynomial decompose() ====== #
    try:
        return decompose(f)
    except ValueError:
        return [f]


def compogen(g_s, symbol):
    """
    Returns the composition of functions.
    Given a list of functions ``g_s``, returns their composition ``f``,
    where:
        f = g_1 o g_2 o .. o g_n

    Note: This is a General composition function. It also composes Polynomials.
    For only Polynomial composition see ``compose`` in polys.

    Examples
    ========

    >>> from sympy.solvers.decompogen import compogen
    >>> from sympy.abc import x
    >>> from sympy import sqrt, sin, cos
    >>> compogen([sin(x), cos(x)], x)
    sin(cos(x))
    >>> compogen([x**2 + x + 1, sin(x)], x)
    sin(x)**2 + sin(x) + 1
    >>> compogen([sqrt(x), 6*x**2 - 5], x)
    sqrt(6*x**2 - 5)
    >>> compogen([sin(x), sqrt(x), cos(x), x**2 + 1], x)
    sin(sqrt(cos(x**2 + 1)))
    >>> compogen([x**2 - x - 1, x**2 + x], x)
    -x**2 - x + (x**2 + x)**2 - 1
    """
    if len(g_s) == 1:
        return g_s[0]

    foo = g_s[0].subs(symbol, g_s[1])

    if len(g_s) == 2:
        return foo

    return compogen([foo] + g_s[2:], symbol)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tools\mypy_annotate.py ===
"""Translates Mypy's output into GitHub's error/warning annotation syntax.

See: https://docs.github.com/en/actions/using-workflows/workflow-commands-for-github-actions

This first is run with Mypy's output piped in, to collect messages in
mypy_annotate.dat. After all platforms run, we run this again, which prints the
messages in GitHub's format but with cross-platform failures deduplicated.
"""

from __future__ import annotations

import argparse
import pickle
import re
import sys

import attrs

# Example: 'package/filename.py:42:1:46:3: error: Type error here [code]'
report_re = re.compile(
    r"""
    ([^:]+):  # Filename (anything but ":")
    ([0-9]+):  # Line number (start)
    (?:([0-9]+):  # Optional column number
      (?:([0-9]+):([0-9]+):)?  # then also optionally, 2 more numbers for end columns
    )?
    \s*(error|warn|note):  # Kind, prefixed with space
    (.+)  # Message
    """,
    re.VERBOSE,
)

mypy_to_github = {
    "error": "error",
    "warn": "warning",
    "note": "notice",
}


@attrs.frozen(kw_only=True)
class Result:
    """Accumulated results, used as a dict key to deduplicate."""

    filename: str
    start_line: int
    kind: str
    message: str
    start_col: int | None = None
    end_line: int | None = None
    end_col: int | None = None


def process_line(line: str) -> Result | None:
    if match := report_re.fullmatch(line.rstrip()):
        filename, st_line, st_col, end_line, end_col, kind, message = match.groups()
        return Result(
            filename=filename,
            start_line=int(st_line),
            start_col=int(st_col) if st_col is not None else None,
            end_line=int(end_line) if end_line is not None else None,
            end_col=int(end_col) if end_col is not None else None,
            kind=mypy_to_github[kind],
            message=message,
        )
    else:
        return None


def export(results: dict[Result, list[str]]) -> None:
    """Display the collected results."""
    for res, platforms in results.items():
        print(f"::{res.kind} file={res.filename},line={res.start_line},", end="")
        if res.start_col is not None:
            print(f"col={res.start_col},", end="")
            if res.end_col is not None and res.end_line is not None:
                print(f"endLine={res.end_line},endColumn={res.end_col},", end="")
                message = f"({res.start_line}:{res.start_col} - {res.end_line}:{res.end_col}):{res.message}"
            else:
                message = f"({res.start_line}:{res.start_col}):{res.message}"
        else:
            message = f"{res.start_line}:{res.message}"
        print(f"title=Mypy-{'+'.join(platforms)}::{res.filename}:{message}")


def main(argv: list[str]) -> None:
    """Look for error messages, and convert the format."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dumpfile",
        help="File to write pickled messages to.",
        required=True,
    )
    parser.add_argument(
        "--platform",
        help="OS name, if set Mypy should be piped to stdin.",
        default=None,
    )
    cmd_line = parser.parse_args(argv)

    results: dict[Result, list[str]]
    try:
        with open(cmd_line.dumpfile, "rb") as f:
            results = pickle.load(f)
    except (FileNotFoundError, pickle.UnpicklingError):
        # If we fail to load, assume it's an old result.
        results = {}

    if cmd_line.platform is None:
        # Write out the results.
        export(results)
    else:
        platform: str = cmd_line.platform
        for line in sys.stdin:
            parsed = process_line(line)
            if parsed is not None:
                try:
                    results[parsed].append(platform)
                except KeyError:
                    results[parsed] = [platform]
            sys.stdout.write(line)
        with open(cmd_line.dumpfile, "wb") as f:
            pickle.dump(results, f)


if __name__ == "__main__":  # pragma: no cover
    main(sys.argv[1:])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\tools\ort_format_model\ort_flatbuffers_py\fbs\Checkpoint.py ===
# automatically generated by the FlatBuffers compiler, do not modify

# namespace: fbs

import flatbuffers
from flatbuffers.compat import import_numpy
np = import_numpy()

class Checkpoint(object):
    __slots__ = ['_tab']

    @classmethod
    def GetRootAs(cls, buf, offset=0):
        n = flatbuffers.encode.Get(flatbuffers.packer.uoffset, buf, offset)
        x = Checkpoint()
        x.Init(buf, n + offset)
        return x

    @classmethod
    def GetRootAsCheckpoint(cls, buf, offset=0):
        """This method is deprecated. Please switch to GetRootAs."""
        return cls.GetRootAs(buf, offset)
    @classmethod
    def CheckpointBufferHasIdentifier(cls, buf, offset, size_prefixed=False):
        return flatbuffers.util.BufferHasIdentifier(buf, offset, b"\x4F\x44\x54\x43", size_prefixed=size_prefixed)

    # Checkpoint
    def Init(self, buf, pos):
        self._tab = flatbuffers.table.Table(buf, pos)

    # Checkpoint
    def Version(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(4))
        if o != 0:
            return self._tab.Get(flatbuffers.number_types.Int32Flags, o + self._tab.Pos)
        return 0

    # Checkpoint
    def ModuleState(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        if o != 0:
            x = self._tab.Indirect(o + self._tab.Pos)
            from ort_flatbuffers_py.fbs.ModuleState import ModuleState
            obj = ModuleState()
            obj.Init(self._tab.Bytes, x)
            return obj
        return None

    # Checkpoint
    def OptimizerGroups(self, j):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(8))
        if o != 0:
            x = self._tab.Vector(o)
            x += flatbuffers.number_types.UOffsetTFlags.py_type(j) * 4
            x = self._tab.Indirect(x)
            from ort_flatbuffers_py.fbs.OptimizerGroup import OptimizerGroup
            obj = OptimizerGroup()
            obj.Init(self._tab.Bytes, x)
            return obj
        return None

    # Checkpoint
    def OptimizerGroupsLength(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(8))
        if o != 0:
            return self._tab.VectorLen(o)
        return 0

    # Checkpoint
    def OptimizerGroupsIsNone(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(8))
        return o == 0

    # Checkpoint
    def PropertyBag(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(10))
        if o != 0:
            x = self._tab.Indirect(o + self._tab.Pos)
            from ort_flatbuffers_py.fbs.PropertyBag import PropertyBag
            obj = PropertyBag()
            obj.Init(self._tab.Bytes, x)
            return obj
        return None

def CheckpointStart(builder):
    builder.StartObject(4)

def Start(builder):
    CheckpointStart(builder)

def CheckpointAddVersion(builder, version):
    builder.PrependInt32Slot(0, version, 0)

def AddVersion(builder, version):
    CheckpointAddVersion(builder, version)

def CheckpointAddModuleState(builder, moduleState):
    builder.PrependUOffsetTRelativeSlot(1, flatbuffers.number_types.UOffsetTFlags.py_type(moduleState), 0)

def AddModuleState(builder, moduleState):
    CheckpointAddModuleState(builder, moduleState)

def CheckpointAddOptimizerGroups(builder, optimizerGroups):
    builder.PrependUOffsetTRelativeSlot(2, flatbuffers.number_types.UOffsetTFlags.py_type(optimizerGroups), 0)

def AddOptimizerGroups(builder, optimizerGroups):
    CheckpointAddOptimizerGroups(builder, optimizerGroups)

def CheckpointStartOptimizerGroupsVector(builder, numElems):
    return builder.StartVector(4, numElems, 4)

def StartOptimizerGroupsVector(builder, numElems: int) -> int:
    return CheckpointStartOptimizerGroupsVector(builder, numElems)

def CheckpointAddPropertyBag(builder, propertyBag):
    builder.PrependUOffsetTRelativeSlot(3, flatbuffers.number_types.UOffsetTFlags.py_type(propertyBag), 0)

def AddPropertyBag(builder, propertyBag):
    CheckpointAddPropertyBag(builder, propertyBag)

def CheckpointEnd(builder):
    return builder.EndObject()

def End(builder):
    return CheckpointEnd(builder)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\_numba\kernels\min_max_.py ===
"""
Numba 1D min/max kernels that can be shared by
* Dataframe / Series
* groupby
* rolling / expanding

Mirrors pandas/_libs/window/aggregation.pyx
"""
from __future__ import annotations

from typing import TYPE_CHECKING

import numba
import numpy as np

if TYPE_CHECKING:
    from pandas._typing import npt


@numba.jit(nopython=True, nogil=True, parallel=False)
def sliding_min_max(
    values: np.ndarray,
    result_dtype: np.dtype,
    start: np.ndarray,
    end: np.ndarray,
    min_periods: int,
    is_max: bool,
) -> tuple[np.ndarray, list[int]]:
    N = len(start)
    nobs = 0
    output = np.empty(N, dtype=result_dtype)
    na_pos = []
    # Use deque once numba supports it
    # https://github.com/numba/numba/issues/7417
    Q: list = []
    W: list = []
    for i in range(N):
        curr_win_size = end[i] - start[i]
        if i == 0:
            st = start[i]
        else:
            st = end[i - 1]

        for k in range(st, end[i]):
            ai = values[k]
            if not np.isnan(ai):
                nobs += 1
            elif is_max:
                ai = -np.inf
            else:
                ai = np.inf
            # Discard previous entries if we find new min or max
            if is_max:
                while Q and ((ai >= values[Q[-1]]) or values[Q[-1]] != values[Q[-1]]):
                    Q.pop()
            else:
                while Q and ((ai <= values[Q[-1]]) or values[Q[-1]] != values[Q[-1]]):
                    Q.pop()
            Q.append(k)
            W.append(k)

        # Discard entries outside and left of current window
        while Q and Q[0] <= start[i] - 1:
            Q.pop(0)
        while W and W[0] <= start[i] - 1:
            if not np.isnan(values[W[0]]):
                nobs -= 1
            W.pop(0)

        # Save output based on index in input value array
        if Q and curr_win_size > 0 and nobs >= min_periods:
            output[i] = values[Q[0]]
        else:
            if values.dtype.kind != "i":
                output[i] = np.nan
            else:
                na_pos.append(i)

    return output, na_pos


@numba.jit(nopython=True, nogil=True, parallel=False)
def grouped_min_max(
    values: np.ndarray,
    result_dtype: np.dtype,
    labels: npt.NDArray[np.intp],
    ngroups: int,
    min_periods: int,
    is_max: bool,
) -> tuple[np.ndarray, list[int]]:
    N = len(labels)
    nobs = np.zeros(ngroups, dtype=np.int64)
    na_pos = []
    output = np.empty(ngroups, dtype=result_dtype)

    for i in range(N):
        lab = labels[i]
        val = values[i]
        if lab < 0:
            continue

        if values.dtype.kind == "i" or not np.isnan(val):
            nobs[lab] += 1
        else:
            # NaN value cannot be a min/max value
            continue

        if nobs[lab] == 1:
            # First element in group, set output equal to this
            output[lab] = val
            continue

        if is_max:
            if val > output[lab]:
                output[lab] = val
        else:
            if val < output[lab]:
                output[lab] = val

    # Set labels that don't satisfy min_periods as np.nan
    for lab, count in enumerate(nobs):
        if count < min_periods:
            na_pos.append(lab)

    return output, na_pos

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\google\protobuf\runtime_version.py ===
# Protocol Buffers - Google's data interchange format
# Copyright 2008 Google Inc.  All rights reserved.
#
# Use of this source code is governed by a BSD-style
# license that can be found in the LICENSE file or at
# https://developers.google.com/open-source/licenses/bsd

"""Protobuf Runtime versions and validators.

It should only be accessed by Protobuf gencodes and tests. DO NOT USE it
elsewhere.
"""

__author__ = 'shaod@google.com (Dennis Shao)'

from enum import Enum
import os
import warnings


class Domain(Enum):
  GOOGLE_INTERNAL = 1
  PUBLIC = 2


# The versions of this Python Protobuf runtime to be changed automatically by
# the Protobuf release process. Do not edit them manually.
# These OSS versions are not stripped to avoid merging conflicts.
OSS_DOMAIN = Domain.PUBLIC
OSS_MAJOR = 6
OSS_MINOR = 31
OSS_PATCH = 1
OSS_SUFFIX = ''

DOMAIN = OSS_DOMAIN
MAJOR = OSS_MAJOR
MINOR = OSS_MINOR
PATCH = OSS_PATCH
SUFFIX = OSS_SUFFIX

# Avoid flooding of warnings.
_MAX_WARNING_COUNT = 20
_warning_count = 0

class VersionError(Exception):
  """Exception class for version violation."""


def _ReportVersionError(msg):
  raise VersionError(msg)


def ValidateProtobufRuntimeVersion(
    gen_domain, gen_major, gen_minor, gen_patch, gen_suffix, location
):
  """Function to validate versions.

  Args:
    gen_domain: The domain where the code was generated from.
    gen_major: The major version number of the gencode.
    gen_minor: The minor version number of the gencode.
    gen_patch: The patch version number of the gencode.
    gen_suffix: The version suffix e.g. '-dev', '-rc1' of the gencode.
    location: The proto location that causes the version violation.

  Raises:
    VersionError: if gencode version is invalid or incompatible with the
    runtime.
  """

  disable_flag = os.getenv('TEMPORARILY_DISABLE_PROTOBUF_VERSION_CHECK')
  if disable_flag is not None and disable_flag.lower() == 'true':
    return

  global _warning_count

  version = f'{MAJOR}.{MINOR}.{PATCH}{SUFFIX}'
  gen_version = f'{gen_major}.{gen_minor}.{gen_patch}{gen_suffix}'

  if gen_major < 0 or gen_minor < 0 or gen_patch < 0:
    raise VersionError(f'Invalid gencode version: {gen_version}')

  error_prompt = (
      'See Protobuf version guarantees at'
      ' https://protobuf.dev/support/cross-version-runtime-guarantee.'
  )

  if gen_domain != DOMAIN:
    _ReportVersionError(
        'Detected mismatched Protobuf Gencode/Runtime domains when loading'
        f' {location}: gencode {gen_domain.name} runtime {DOMAIN.name}.'
        ' Cross-domain usage of Protobuf is not supported.'
    )

  if gen_major != MAJOR:
    if gen_major == MAJOR - 1:
      if _warning_count < _MAX_WARNING_COUNT:
        warnings.warn(
            'Protobuf gencode version %s is exactly one major version older'
            ' than the runtime version %s at %s. Please update the gencode to'
            ' avoid compatibility violations in the next runtime release.'
            % (gen_version, version, location)
        )
        _warning_count += 1
    else:
      _ReportVersionError(
          'Detected mismatched Protobuf Gencode/Runtime major versions when'
          f' loading {location}: gencode {gen_version} runtime {version}.'
          f' Same major version is required. {error_prompt}'
      )

  if MINOR < gen_minor or (MINOR == gen_minor and PATCH < gen_patch):
    _ReportVersionError(
        'Detected incompatible Protobuf Gencode/Runtime versions when loading'
        f' {location}: gencode {gen_version} runtime {version}. Runtime version'
        f' cannot be older than the linked gencode version. {error_prompt}'
    )

  if gen_suffix != SUFFIX:
    _ReportVersionError(
        'Detected mismatched Protobuf Gencode/Runtime version suffixes when'
        f' loading {location}: gencode {gen_version} runtime {version}.'
        f' Version suffixes must be the same. {error_prompt}'
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\chart\reference.py ===
# Copyright (c) 2010-2024 openpyxl

from itertools import chain

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    MinMax,
    Typed,
    String,
    Strict,
)
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import (
    get_column_letter,
    range_to_tuple,
    quote_sheetname
)


class DummyWorksheet:


    def __init__(self, title):
        self.title = title


class Reference(Strict):

    """
    Normalise cell range references
    """

    min_row = MinMax(min=1, max=1000000, expected_type=int)
    max_row = MinMax(min=1, max=1000000, expected_type=int)
    min_col = MinMax(min=1, max=16384, expected_type=int)
    max_col = MinMax(min=1, max=16384, expected_type=int)
    range_string = String(allow_none=True)

    def __init__(self,
                 worksheet=None,
                 min_col=None,
                 min_row=None,
                 max_col=None,
                 max_row=None,
                 range_string=None
                 ):
        if range_string is not None:
            sheetname, boundaries = range_to_tuple(range_string)
            min_col, min_row, max_col, max_row = boundaries
            worksheet = DummyWorksheet(sheetname)

        self.worksheet = worksheet
        self.min_col = min_col
        self.min_row = min_row
        if max_col is None:
            max_col = min_col
        self.max_col = max_col
        if max_row is None:
            max_row = min_row
        self.max_row = max_row


    def __repr__(self):
        return str(self)


    def __str__(self):
        fmt = u"{0}!${1}${2}:${3}${4}"
        if (self.min_col == self.max_col
            and self.min_row == self.max_row):
            fmt = u"{0}!${1}${2}"
        return fmt.format(self.sheetname,
                          get_column_letter(self.min_col), self.min_row,
                          get_column_letter(self.max_col), self.max_row
                          )


    __str__ = __str__



    def __len__(self):
        if self.min_row == self.max_row:
            return 1 + self.max_col - self.min_col
        return 1 + self.max_row - self.min_row


    def __eq__(self, other):
        return str(self) == str(other)


    @property
    def rows(self):
        """
        Return all rows in the range
        """
        for row in range(self.min_row, self.max_row+1):
            yield Reference(self.worksheet, self.min_col, row, self.max_col, row)


    @property
    def cols(self):
        """
        Return all columns in the range
        """
        for col in range(self.min_col, self.max_col+1):
            yield Reference(self.worksheet, col, self.min_row, col, self.max_row)


    def pop(self):
        """
        Return and remove the first cell
        """
        cell = "{0}{1}".format(get_column_letter(self.min_col), self.min_row)
        if self.min_row == self.max_row:
            self.min_col += 1
        else:
            self.min_row += 1
        return cell


    @property
    def sheetname(self):
        return quote_sheetname(self.worksheet.title)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\utils\deprecation.py ===
"""
A module that implements tooling to enable easy warnings about deprecations.
"""

import logging
import warnings
from typing import Any, Optional, TextIO, Type, Union

from pip._vendor.packaging.version import parse

from pip import __version__ as current_version  # NOTE: tests patch this name.

DEPRECATION_MSG_PREFIX = "DEPRECATION: "


class PipDeprecationWarning(Warning):
    pass


_original_showwarning: Any = None


# Warnings <-> Logging Integration
def _showwarning(
    message: Union[Warning, str],
    category: Type[Warning],
    filename: str,
    lineno: int,
    file: Optional[TextIO] = None,
    line: Optional[str] = None,
) -> None:
    if file is not None:
        if _original_showwarning is not None:
            _original_showwarning(message, category, filename, lineno, file, line)
    elif issubclass(category, PipDeprecationWarning):
        # We use a specially named logger which will handle all of the
        # deprecation messages for pip.
        logger = logging.getLogger("pip._internal.deprecations")
        logger.warning(message)
    else:
        _original_showwarning(message, category, filename, lineno, file, line)


def install_warning_logger() -> None:
    # Enable our Deprecation Warnings
    warnings.simplefilter("default", PipDeprecationWarning, append=True)

    global _original_showwarning

    if _original_showwarning is None:
        _original_showwarning = warnings.showwarning
        warnings.showwarning = _showwarning


def deprecated(
    *,
    reason: str,
    replacement: Optional[str],
    gone_in: Optional[str],
    feature_flag: Optional[str] = None,
    issue: Optional[int] = None,
) -> None:
    """Helper to deprecate existing functionality.

    reason:
        Textual reason shown to the user about why this functionality has
        been deprecated. Should be a complete sentence.
    replacement:
        Textual suggestion shown to the user about what alternative
        functionality they can use.
    gone_in:
        The version of pip does this functionality should get removed in.
        Raises an error if pip's current version is greater than or equal to
        this.
    feature_flag:
        Command-line flag of the form --use-feature={feature_flag} for testing
        upcoming functionality.
    issue:
        Issue number on the tracker that would serve as a useful place for
        users to find related discussion and provide feedback.
    """

    # Determine whether or not the feature is already gone in this version.
    is_gone = gone_in is not None and parse(current_version) >= parse(gone_in)

    message_parts = [
        (reason, f"{DEPRECATION_MSG_PREFIX}{{}}"),
        (
            gone_in,
            (
                "pip {} will enforce this behaviour change."
                if not is_gone
                else "Since pip {}, this is no longer supported."
            ),
        ),
        (
            replacement,
            "A possible replacement is {}.",
        ),
        (
            feature_flag,
            (
                "You can use the flag --use-feature={} to test the upcoming behaviour."
                if not is_gone
                else None
            ),
        ),
        (
            issue,
            "Discussion can be found at https://github.com/pypa/pip/issues/{}",
        ),
    ]

    message = " ".join(
        format_str.format(value)
        for value, format_str in message_parts
        if format_str is not None and value is not None
    )

    # Raise as an error if this behaviour is deprecated.
    if is_gone:
        raise PipDeprecationWarning(message)

    warnings.warn(message, category=PipDeprecationWarning, stacklevel=2)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pyreadline3\lineeditor\wordmatcher.py ===
# -*- coding: utf-8 -*-
# *****************************************************************************
#       Copyright (C) 2006-2020 Jorgen Stenarson. <jorgen.stenarson@bostream.nu>
#       Copyright (C) 2020 Bassem Girgis. <brgirgis@gmail.com>
#
#  Distributed under the terms of the BSD License.  The full license is in
#  the file COPYING, distributed as part of this software.
# *****************************************************************************


import re


def str_find_all(in_str, ch):
    result = []
    index = 0
    while index >= 0:
        index = in_str.find(ch, index)
        if index >= 0:
            result.append(index)
            index += 1
    return result


word_pattern = re.compile("(x*)")


def markwords(in_str, is_wordfun):
    markers = {True: "x", False: "o"}
    return "".join([markers[is_wordfun(ch)] for ch in in_str])


def split_words(in_str, is_wordfun):
    return [x for x in word_pattern.split(markwords(in_str, is_wordfun)) if x != ""]


def mark_start_segment(in_str, is_segment):
    def mark_start(s):
        if s[0:1] == "x":
            return "s" + s[1:]
        else:
            return s

    return "".join(map(mark_start, split_words(in_str, is_segment)))


def mark_end_segment(in_str, is_segment):
    def mark_start(s):
        if s[0:1] == "x":
            return s[:-1] + "s"
        else:
            return s

    return "".join(map(mark_start, split_words(in_str, is_segment)))


def mark_start_segment_index(in_str, is_segment):
    return str_find_all(mark_start_segment(in_str, is_segment), "s")


def mark_end_segment_index(in_str, is_segment):
    return [x + 1 for x in str_find_all(mark_end_segment(in_str, is_segment), "s")]


# ###############  Following are used in lineobj  ###########################


def is_word_token(in_str):
    return not is_non_word_token(in_str)


def is_non_word_token(in_str):
    if len(in_str) != 1 or in_str in " \t\n":
        return True
    else:
        return False


def next_start_segment(in_str, is_segment):
    in_str = "".join(in_str)
    result = []
    for start in mark_start_segment_index(in_str, is_segment):
        result[len(result) : start] = [start for x in range(start - len(result))]
    result[len(result) : len(in_str)] = [
        len(in_str) for x in range(len(in_str) - len(result) + 1)
    ]
    return result


def next_end_segment(in_str, is_segment):
    in_str = "".join(in_str)
    result = []
    for start in mark_end_segment_index(in_str, is_segment):
        result[len(result) : start] = [start for x in range(start - len(result))]
    result[len(result) : len(in_str)] = [
        len(in_str) for x in range(len(in_str) - len(result) + 1)
    ]
    return result


def prev_start_segment(in_str, is_segment):
    in_str = "".join(in_str)
    result = []
    prev = 0
    for start in mark_start_segment_index(in_str, is_segment):
        result[len(result) : start + 1] = [prev for x in range(start - len(result) + 1)]
        prev = start
    result[len(result) : len(in_str)] = [
        prev for x in range(len(in_str) - len(result) + 1)
    ]
    return result


def prev_end_segment(in_str, is_segment):
    in_str = "".join(in_str)
    result = []
    prev = 0
    for start in mark_end_segment_index(in_str, is_segment):
        result[len(result) : start + 1] = [prev for x in range(start - len(result) + 1)]
        prev = start
    result[len(result) : len(in_str)] = [
        len(in_str) for x in range(len(in_str) - len(result) + 1)
    ]
    return result

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\plotting\pygletplot\plot_camera.py ===
import pyglet.gl as pgl
from sympy.plotting.pygletplot.plot_rotation import get_spherical_rotatation
from sympy.plotting.pygletplot.util import get_model_matrix, model_to_screen, \
                                            screen_to_model, vec_subs


class PlotCamera:

    min_dist = 0.05
    max_dist = 500.0

    min_ortho_dist = 100.0
    max_ortho_dist = 10000.0

    _default_dist = 6.0
    _default_ortho_dist = 600.0

    rot_presets = {
        'xy': (0, 0, 0),
        'xz': (-90, 0, 0),
        'yz': (0, 90, 0),
        'perspective': (-45, 0, -45)
    }

    def __init__(self, window, ortho=False):
        self.window = window
        self.axes = self.window.plot.axes
        self.ortho = ortho
        self.reset()

    def init_rot_matrix(self):
        pgl.glPushMatrix()
        pgl.glLoadIdentity()
        self._rot = get_model_matrix()
        pgl.glPopMatrix()

    def set_rot_preset(self, preset_name):
        self.init_rot_matrix()
        if preset_name not in self.rot_presets:
            raise ValueError(
                "%s is not a valid rotation preset." % preset_name)
        r = self.rot_presets[preset_name]
        self.euler_rotate(r[0], 1, 0, 0)
        self.euler_rotate(r[1], 0, 1, 0)
        self.euler_rotate(r[2], 0, 0, 1)

    def reset(self):
        self._dist = 0.0
        self._x, self._y = 0.0, 0.0
        self._rot = None
        if self.ortho:
            self._dist = self._default_ortho_dist
        else:
            self._dist = self._default_dist
        self.init_rot_matrix()

    def mult_rot_matrix(self, rot):
        pgl.glPushMatrix()
        pgl.glLoadMatrixf(rot)
        pgl.glMultMatrixf(self._rot)
        self._rot = get_model_matrix()
        pgl.glPopMatrix()

    def setup_projection(self):
        pgl.glMatrixMode(pgl.GL_PROJECTION)
        pgl.glLoadIdentity()
        if self.ortho:
            # yep, this is pseudo ortho (don't tell anyone)
            pgl.gluPerspective(
                0.3, float(self.window.width)/float(self.window.height),
                self.min_ortho_dist - 0.01, self.max_ortho_dist + 0.01)
        else:
            pgl.gluPerspective(
                30.0, float(self.window.width)/float(self.window.height),
                self.min_dist - 0.01, self.max_dist + 0.01)
        pgl.glMatrixMode(pgl.GL_MODELVIEW)

    def _get_scale(self):
        return 1.0, 1.0, 1.0

    def apply_transformation(self):
        pgl.glLoadIdentity()
        pgl.glTranslatef(self._x, self._y, -self._dist)
        if self._rot is not None:
            pgl.glMultMatrixf(self._rot)
        pgl.glScalef(*self._get_scale())

    def spherical_rotate(self, p1, p2, sensitivity=1.0):
        mat = get_spherical_rotatation(p1, p2, self.window.width,
                                       self.window.height, sensitivity)
        if mat is not None:
            self.mult_rot_matrix(mat)

    def euler_rotate(self, angle, x, y, z):
        pgl.glPushMatrix()
        pgl.glLoadMatrixf(self._rot)
        pgl.glRotatef(angle, x, y, z)
        self._rot = get_model_matrix()
        pgl.glPopMatrix()

    def zoom_relative(self, clicks, sensitivity):

        if self.ortho:
            dist_d = clicks * sensitivity * 50.0
            min_dist = self.min_ortho_dist
            max_dist = self.max_ortho_dist
        else:
            dist_d = clicks * sensitivity
            min_dist = self.min_dist
            max_dist = self.max_dist

        new_dist = (self._dist - dist_d)
        if (clicks < 0 and new_dist < max_dist) or new_dist > min_dist:
            self._dist = new_dist

    def mouse_translate(self, x, y, dx, dy):
        pgl.glPushMatrix()
        pgl.glLoadIdentity()
        pgl.glTranslatef(0, 0, -self._dist)
        z = model_to_screen(0, 0, 0)[2]
        d = vec_subs(screen_to_model(x, y, z), screen_to_model(x - dx, y - dy, z))
        pgl.glPopMatrix()
        self._x += d[0]
        self._y += d[1]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\unify\usympy.py ===
""" SymPy interface to Unification engine

See sympy.unify for module level docstring
See sympy.unify.core for algorithmic docstring """

from sympy.core import Basic, Add, Mul, Pow
from sympy.core.operations import AssocOp, LatticeOp
from sympy.matrices import MatAdd, MatMul, MatrixExpr
from sympy.sets.sets import Union, Intersection, FiniteSet
from sympy.unify.core import Compound, Variable, CondVariable
from sympy.unify import core

basic_new_legal = [MatrixExpr]
eval_false_legal = [AssocOp, Pow, FiniteSet]
illegal = [LatticeOp]

def sympy_associative(op):
    assoc_ops = (AssocOp, MatAdd, MatMul, Union, Intersection, FiniteSet)
    return any(issubclass(op, aop) for aop in assoc_ops)

def sympy_commutative(op):
    comm_ops = (Add, MatAdd, Union, Intersection, FiniteSet)
    return any(issubclass(op, cop) for cop in comm_ops)

def is_associative(x):
    return isinstance(x, Compound) and sympy_associative(x.op)

def is_commutative(x):
    if not isinstance(x, Compound):
        return False
    if sympy_commutative(x.op):
        return True
    if issubclass(x.op, Mul):
        return all(construct(arg).is_commutative for arg in x.args)

def mk_matchtype(typ):
    def matchtype(x):
        return (isinstance(x, typ) or
                isinstance(x, Compound) and issubclass(x.op, typ))
    return matchtype

def deconstruct(s, variables=()):
    """ Turn a SymPy object into a Compound """
    if s in variables:
        return Variable(s)
    if isinstance(s, (Variable, CondVariable)):
        return s
    if not isinstance(s, Basic) or s.is_Atom:
        return s
    return Compound(s.__class__,
                    tuple(deconstruct(arg, variables) for arg in s.args))

def construct(t):
    """ Turn a Compound into a SymPy object """
    if isinstance(t, (Variable, CondVariable)):
        return t.arg
    if not isinstance(t, Compound):
        return t
    if any(issubclass(t.op, cls) for cls in eval_false_legal):
        return t.op(*map(construct, t.args), evaluate=False)
    elif any(issubclass(t.op, cls) for cls in basic_new_legal):
        return Basic.__new__(t.op, *map(construct, t.args))
    else:
        return t.op(*map(construct, t.args))

def rebuild(s):
    """ Rebuild a SymPy expression.

    This removes harm caused by Expr-Rules interactions.
    """
    return construct(deconstruct(s))

def unify(x, y, s=None, variables=(), **kwargs):
    """ Structural unification of two expressions/patterns.

    Examples
    ========

    >>> from sympy.unify.usympy import unify
    >>> from sympy import Basic, S
    >>> from sympy.abc import x, y, z, p, q

    >>> next(unify(Basic(S(1), S(2)), Basic(S(1), x), variables=[x]))
    {x: 2}

    >>> expr = 2*x + y + z
    >>> pattern = 2*p + q
    >>> next(unify(expr, pattern, {}, variables=(p, q)))
    {p: x, q: y + z}

    Unification supports commutative and associative matching

    >>> expr = x + y + z
    >>> pattern = p + q
    >>> len(list(unify(expr, pattern, {}, variables=(p, q))))
    12

    Symbols not indicated to be variables are treated as literal,
    else they are wild-like and match anything in a sub-expression.

    >>> expr = x*y*z + 3
    >>> pattern = x*y + 3
    >>> next(unify(expr, pattern, {}, variables=[x, y]))
    {x: y, y: x*z}

    The x and y of the pattern above were in a Mul and matched factors
    in the Mul of expr. Here, a single symbol matches an entire term:

    >>> expr = x*y + 3
    >>> pattern = p + 3
    >>> next(unify(expr, pattern, {}, variables=[p]))
    {p: x*y}

    """
    decons = lambda x: deconstruct(x, variables)
    s = s or {}
    s = {decons(k): decons(v) for k, v in s.items()}

    ds = core.unify(decons(x), decons(y), s,
                                     is_associative=is_associative,
                                     is_commutative=is_commutative,
                                     **kwargs)
    for d in ds:
        yield {construct(k): construct(v) for k, v in d.items()}

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\urllib3\util\wait.py ===
from __future__ import annotations

import select
import socket
from functools import partial

__all__ = ["wait_for_read", "wait_for_write"]


# How should we wait on sockets?
#
# There are two types of APIs you can use for waiting on sockets: the fancy
# modern stateful APIs like epoll/kqueue, and the older stateless APIs like
# select/poll. The stateful APIs are more efficient when you have a lots of
# sockets to keep track of, because you can set them up once and then use them
# lots of times. But we only ever want to wait on a single socket at a time
# and don't want to keep track of state, so the stateless APIs are actually
# more efficient. So we want to use select() or poll().
#
# Now, how do we choose between select() and poll()? On traditional Unixes,
# select() has a strange calling convention that makes it slow, or fail
# altogether, for high-numbered file descriptors. The point of poll() is to fix
# that, so on Unixes, we prefer poll().
#
# On Windows, there is no poll() (or at least Python doesn't provide a wrapper
# for it), but that's OK, because on Windows, select() doesn't have this
# strange calling convention; plain select() works fine.
#
# So: on Windows we use select(), and everywhere else we use poll(). We also
# fall back to select() in case poll() is somehow broken or missing.


def select_wait_for_socket(
    sock: socket.socket,
    read: bool = False,
    write: bool = False,
    timeout: float | None = None,
) -> bool:
    if not read and not write:
        raise RuntimeError("must specify at least one of read=True, write=True")
    rcheck = []
    wcheck = []
    if read:
        rcheck.append(sock)
    if write:
        wcheck.append(sock)
    # When doing a non-blocking connect, most systems signal success by
    # marking the socket writable. Windows, though, signals success by marked
    # it as "exceptional". We paper over the difference by checking the write
    # sockets for both conditions. (The stdlib selectors module does the same
    # thing.)
    fn = partial(select.select, rcheck, wcheck, wcheck)
    rready, wready, xready = fn(timeout)
    return bool(rready or wready or xready)


def poll_wait_for_socket(
    sock: socket.socket,
    read: bool = False,
    write: bool = False,
    timeout: float | None = None,
) -> bool:
    if not read and not write:
        raise RuntimeError("must specify at least one of read=True, write=True")
    mask = 0
    if read:
        mask |= select.POLLIN
    if write:
        mask |= select.POLLOUT
    poll_obj = select.poll()
    poll_obj.register(sock, mask)

    # For some reason, poll() takes timeout in milliseconds
    def do_poll(t: float | None) -> list[tuple[int, int]]:
        if t is not None:
            t *= 1000
        return poll_obj.poll(t)

    return bool(do_poll(timeout))


def _have_working_poll() -> bool:
    # Apparently some systems have a select.poll that fails as soon as you try
    # to use it, either due to strange configuration or broken monkeypatching
    # from libraries like eventlet/greenlet.
    try:
        poll_obj = select.poll()
        poll_obj.poll(0)
    except (AttributeError, OSError):
        return False
    else:
        return True


def wait_for_socket(
    sock: socket.socket,
    read: bool = False,
    write: bool = False,
    timeout: float | None = None,
) -> bool:
    # We delay choosing which implementation to use until the first time we're
    # called. We could do it at import time, but then we might make the wrong
    # decision if someone goes wild with monkeypatching select.poll after
    # we're imported.
    global wait_for_socket
    if _have_working_poll():
        wait_for_socket = poll_wait_for_socket
    elif hasattr(select, "select"):
        wait_for_socket = select_wait_for_socket
    return wait_for_socket(sock, read, write, timeout)


def wait_for_read(sock: socket.socket, timeout: float | None = None) -> bool:
    """Waits for reading to be available on a given socket.
    Returns True if the socket is readable, or False if the timeout expired.
    """
    return wait_for_socket(sock, read=True, timeout=timeout)


def wait_for_write(sock: socket.socket, timeout: float | None = None) -> bool:
    """Waits for writing to be available on a given socket.
    Returns True if the socket is readable, or False if the timeout expired.
    """
    return wait_for_socket(sock, write=True, timeout=timeout)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\tests\test_polytools.py ===
"""Tests for user-friendly public interface to polynomial functions. """

import pickle

from sympy.polys.polytools import (
    Poly, PurePoly, poly,
    parallel_poly_from_expr,
    degree, degree_list,
    total_degree,
    LC, LM, LT,
    pdiv, prem, pquo, pexquo,
    div, rem, quo, exquo,
    half_gcdex, gcdex, invert,
    subresultants,
    resultant, discriminant,
    terms_gcd, cofactors,
    gcd, gcd_list,
    lcm, lcm_list,
    trunc,
    monic, content, primitive,
    compose, decompose,
    sturm,
    gff_list, gff,
    sqf_norm, sqf_part, sqf_list, sqf,
    factor_list, factor,
    intervals, refine_root, count_roots,
    all_roots, real_roots, nroots, ground_roots,
    nth_power_roots_poly,
    cancel, reduced, groebner,
    GroebnerBasis, is_zero_dimensional,
    _torational_factor_list,
    to_rational_coeffs)

from sympy.polys.polyerrors import (
    MultivariatePolynomialError,
    ExactQuotientFailed,
    PolificationFailed,
    ComputationFailed,
    UnificationFailed,
    RefinementFailed,
    GeneratorsNeeded,
    GeneratorsError,
    PolynomialError,
    CoercionFailed,
    DomainError,
    OptionError,
    FlagError)

from sympy.polys.polyclasses import DMP

from sympy.polys.fields import field
from sympy.polys.domains import FF, ZZ, QQ, ZZ_I, QQ_I, RR, EX
from sympy.polys.domains.realfield import RealField
from sympy.polys.domains.complexfield import ComplexField
from sympy.polys.orderings import lex, grlex, grevlex

from sympy.combinatorics.galois import S4TransitiveSubgroups
from sympy.core.add import Add
from sympy.core.basic import _aresame
from sympy.core.containers import Tuple
from sympy.core.expr import Expr
from sympy.core.function import (Derivative, diff, expand)
from sympy.core.mul import _keep_coeff, Mul
from sympy.core.numbers import (Float, I, Integer, Rational, oo, pi)
from sympy.core.power import Pow
from sympy.core.relational import Eq
from sympy.core.singleton import S
from sympy.core.symbol import Symbol
from sympy.functions.elementary.complexes import (im, re)
from sympy.functions.elementary.exponential import exp
from sympy.functions.elementary.hyperbolic import tanh
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.piecewise import Piecewise
from sympy.functions.elementary.trigonometric import sin
from sympy.matrices.dense import Matrix
from sympy.matrices.expressions.matexpr import MatrixSymbol
from sympy.polys.rootoftools import rootof
from sympy.simplify.simplify import signsimp
from sympy.utilities.iterables import iterable
from sympy.utilities.exceptions import SymPyDeprecationWarning

from sympy.testing.pytest import (
    raises, warns_deprecated_sympy, warns, tooslow, XFAIL
)

from sympy.abc import a, b, c, d, p, q, t, w, x, y, z


def _epsilon_eq(a, b):
    for u, v in zip(a, b):
        if abs(u - v) > 1e-10:
            return False
    return True


def _strict_eq(a, b):
    if type(a) == type(b):
        if iterable(a):
            if len(a) == len(b):
                return all(_strict_eq(c, d) for c, d in zip(a, b))
            else:
                return False
        else:
            return isinstance(a, Poly) and a.eq(b, strict=True)
    else:
        return False


def test_Poly_mixed_operations():
    p = Poly(x, x)
    with warns_deprecated_sympy():
        p * exp(x)
    with warns_deprecated_sympy():
        p + exp(x)
    with warns_deprecated_sympy():
        p - exp(x)


def test_Poly_from_dict():
    K = FF(3)

    assert Poly.from_dict(
        {0: 1, 1: 2}, gens=x, domain=K).rep == DMP([K(2), K(1)], K)
    assert Poly.from_dict(
        {0: 1, 1: 5}, gens=x, domain=K).rep == DMP([K(2), K(1)], K)

    assert Poly.from_dict(
        {(0,): 1, (1,): 2}, gens=x, domain=K).rep == DMP([K(2), K(1)], K)
    assert Poly.from_dict(
        {(0,): 1, (1,): 5}, gens=x, domain=K).rep == DMP([K(2), K(1)], K)

    assert Poly.from_dict({(0, 0): 1, (1, 1): 2}, gens=(
        x, y), domain=K).rep == DMP([[K(2), K(0)], [K(1)]], K)

    assert Poly.from_dict({0: 1, 1: 2}, gens=x).rep == DMP([ZZ(2), ZZ(1)], ZZ)
    assert Poly.from_dict(
        {0: 1, 1: 2}, gens=x, field=True).rep == DMP([QQ(2), QQ(1)], QQ)

    assert Poly.from_dict(
        {0: 1, 1: 2}, gens=x, domain=ZZ).rep == DMP([ZZ(2), ZZ(1)], ZZ)
    assert Poly.from_dict(
        {0: 1, 1: 2}, gens=x, domain=QQ).rep == DMP([QQ(2), QQ(1)], QQ)

    assert Poly.from_dict(
        {(0,): 1, (1,): 2}, gens=x).rep == DMP([ZZ(2), ZZ(1)], ZZ)
    assert Poly.from_dict(
        {(0,): 1, (1,): 2}, gens=x, field=True).rep == DMP([QQ(2), QQ(1)], QQ)

    assert Poly.from_dict(
        {(0,): 1, (1,): 2}, gens=x, domain=ZZ).rep == DMP([ZZ(2), ZZ(1)], ZZ)
    assert Poly.from_dict(
        {(0,): 1, (1,): 2}, gens=x, domain=QQ).rep == DMP([QQ(2), QQ(1)], QQ)

    assert Poly.from_dict({(1,): sin(y)}, gens=x, composite=False) == \
        Poly(sin(y)*x, x, domain='EX')
    assert Poly.from_dict({(1,): y}, gens=x, composite=False) == \
        Poly(y*x, x, domain='EX')
    assert Poly.from_dict({(1, 1): 1}, gens=(x, y), composite=False) == \
        Poly(x*y, x, y, domain='ZZ')
    assert Poly.from_dict({(1, 0): y}, gens=(x, z), composite=False) == \
        Poly(y*x, x, z, domain='EX')


def test_Poly_from_list():
    K = FF(3)

    assert Poly.from_list([2, 1], gens=x, domain=K).rep == DMP([K(2), K(1)], K)
    assert Poly.from_list([5, 1], gens=x, domain=K).rep == DMP([K(2), K(1)], K)

    assert Poly.from_list([2, 1], gens=x).rep == DMP([ZZ(2), ZZ(1)], ZZ)
    assert Poly.from_list([2, 1], gens=x, field=True).rep == DMP([QQ(2), QQ(1)], QQ)

    assert Poly.from_list([2, 1], gens=x, domain=ZZ).rep == DMP([ZZ(2), ZZ(1)], ZZ)
    assert Poly.from_list([2, 1], gens=x, domain=QQ).rep == DMP([QQ(2), QQ(1)], QQ)

    assert Poly.from_list([0, 1.0], gens=x).rep == DMP([RR(1.0)], RR)
    assert Poly.from_list([1.0, 0], gens=x).rep == DMP([RR(1.0), RR(0.0)], RR)

    raises(MultivariatePolynomialError, lambda: Poly.from_list([[]], gens=(x, y)))


def test_Poly_from_poly():
    f = Poly(x + 7, x, domain=ZZ)
    g = Poly(x + 2, x, modulus=3)
    h = Poly(x + y, x, y, domain=ZZ)

    K = FF(3)

    assert Poly.from_poly(f) == f
    assert Poly.from_poly(f, domain=K).rep == DMP([K(1), K(1)], K)
    assert Poly.from_poly(f, domain=ZZ).rep == DMP([ZZ(1), ZZ(7)], ZZ)
    assert Poly.from_poly(f, domain=QQ).rep == DMP([QQ(1), QQ(7)], QQ)

    assert Poly.from_poly(f, gens=x) == f
    assert Poly.from_poly(f, gens=x, domain=K).rep == DMP([K(1), K(1)], K)
    assert Poly.from_poly(f, gens=x, domain=ZZ).rep == DMP([ZZ(1), ZZ(7)], ZZ)
    assert Poly.from_poly(f, gens=x, domain=QQ).rep == DMP([QQ(1), QQ(7)], QQ)

    assert Poly.from_poly(f, gens=y) == Poly(x + 7, y, domain='ZZ[x]')
    raises(CoercionFailed, lambda: Poly.from_poly(f, gens=y, domain=K))
    raises(CoercionFailed, lambda: Poly.from_poly(f, gens=y, domain=ZZ))
    raises(CoercionFailed, lambda: Poly.from_poly(f, gens=y, domain=QQ))

    assert Poly.from_poly(f, gens=(x, y)) == Poly(x + 7, x, y, domain='ZZ')
    assert Poly.from_poly(
        f, gens=(x, y), domain=ZZ) == Poly(x + 7, x, y, domain='ZZ')
    assert Poly.from_poly(
        f, gens=(x, y), domain=QQ) == Poly(x + 7, x, y, domain='QQ')
    assert Poly.from_poly(
        f, gens=(x, y), modulus=3) == Poly(x + 7, x, y, domain='FF(3)')

    K = FF(2)

    assert Poly.from_poly(g) == g
    assert Poly.from_poly(g, domain=ZZ).rep == DMP([ZZ(1), ZZ(-1)], ZZ)
    raises(CoercionFailed, lambda: Poly.from_poly(g, domain=QQ))
    assert Poly.from_poly(g, domain=K).rep == DMP([K(1), K(0)], K)

    assert Poly.from_poly(g, gens=x) == g
    assert Poly.from_poly(g, gens=x, domain=ZZ).rep == DMP([ZZ(1), ZZ(-1)], ZZ)
    raises(CoercionFailed, lambda: Poly.from_poly(g, gens=x, domain=QQ))
    assert Poly.from_poly(g, gens=x, domain=K).rep == DMP([K(1), K(0)], K)

    K = FF(3)

    assert Poly.from_poly(h) == h
    assert Poly.from_poly(
        h, domain=ZZ).rep == DMP([[ZZ(1)], [ZZ(1), ZZ(0)]], ZZ)
    assert Poly.from_poly(
        h, domain=QQ).rep == DMP([[QQ(1)], [QQ(1), QQ(0)]], QQ)
    assert Poly.from_poly(h, domain=K).rep == DMP([[K(1)], [K(1), K(0)]], K)

    assert Poly.from_poly(h, gens=x) == Poly(x + y, x, domain=ZZ[y])
    raises(CoercionFailed, lambda: Poly.from_poly(h, gens=x, domain=ZZ))
    assert Poly.from_poly(
        h, gens=x, domain=ZZ[y]) == Poly(x + y, x, domain=ZZ[y])
    raises(CoercionFailed, lambda: Poly.from_poly(h, gens=x, domain=QQ))
    assert Poly.from_poly(
        h, gens=x, domain=QQ[y]) == Poly(x + y, x, domain=QQ[y])
    raises(CoercionFailed, lambda: Poly.from_poly(h, gens=x, modulus=3))

    assert Poly.from_poly(h, gens=y) == Poly(x + y, y, domain=ZZ[x])
    raises(CoercionFailed, lambda: Poly.from_poly(h, gens=y, domain=ZZ))
    assert Poly.from_poly(
        h, gens=y, domain=ZZ[x]) == Poly(x + y, y, domain=ZZ[x])
    raises(CoercionFailed, lambda: Poly.from_poly(h, gens=y, domain=QQ))
    assert Poly.from_poly(
        h, gens=y, domain=QQ[x]) == Poly(x + y, y, domain=QQ[x])
    raises(CoercionFailed, lambda: Poly.from_poly(h, gens=y, modulus=3))

    assert Poly.from_poly(h, gens=(x, y)) == h
    assert Poly.from_poly(
        h, gens=(x, y), domain=ZZ).rep == DMP([[ZZ(1)], [ZZ(1), ZZ(0)]], ZZ)
    assert Poly.from_poly(
        h, gens=(x, y), domain=QQ).rep == DMP([[QQ(1)], [QQ(1), QQ(0)]], QQ)
    assert Poly.from_poly(
        h, gens=(x, y), domain=K).rep == DMP([[K(1)], [K(1), K(0)]], K)

    assert Poly.from_poly(
        h, gens=(y, x)).rep == DMP([[ZZ(1)], [ZZ(1), ZZ(0)]], ZZ)
    assert Poly.from_poly(
        h, gens=(y, x), domain=ZZ).rep == DMP([[ZZ(1)], [ZZ(1), ZZ(0)]], ZZ)
    assert Poly.from_poly(
        h, gens=(y, x), domain=QQ).rep == DMP([[QQ(1)], [QQ(1), QQ(0)]], QQ)
    assert Poly.from_poly(
        h, gens=(y, x), domain=K).rep == DMP([[K(1)], [K(1), K(0)]], K)

    assert Poly.from_poly(
        h, gens=(x, y), field=True).rep == DMP([[QQ(1)], [QQ(1), QQ(0)]], QQ)
    assert Poly.from_poly(
        h, gens=(x, y), field=True).rep == DMP([[QQ(1)], [QQ(1), QQ(0)]], QQ)


def test_Poly_from_expr():
    raises(GeneratorsNeeded, lambda: Poly.from_expr(S.Zero))
    raises(GeneratorsNeeded, lambda: Poly.from_expr(S(7)))

    F3 = FF(3)

    assert Poly.from_expr(x + 5, domain=F3).rep == DMP([F3(1), F3(2)], F3)
    assert Poly.from_expr(y + 5, domain=F3).rep == DMP([F3(1), F3(2)], F3)

    assert Poly.from_expr(x + 5, x, domain=F3).rep == DMP([F3(1), F3(2)], F3)
    assert Poly.from_expr(y + 5, y, domain=F3).rep == DMP([F3(1), F3(2)], F3)

    assert Poly.from_expr(x + y, domain=F3).rep == DMP([[F3(1)], [F3(1), F3(0)]], F3)
    assert Poly.from_expr(x + y, x, y, domain=F3).rep == DMP([[F3(1)], [F3(1), F3(0)]], F3)

    assert Poly.from_expr(x + 5).rep == DMP([ZZ(1), ZZ(5)], ZZ)
    assert Poly.from_expr(y + 5).rep == DMP([ZZ(1), ZZ(5)], ZZ)

    assert Poly.from_expr(x + 5, x).rep == DMP([ZZ(1), ZZ(5)], ZZ)
    assert Poly.from_expr(y + 5, y).rep == DMP([ZZ(1), ZZ(5)], ZZ)

    assert Poly.from_expr(x + 5, domain=ZZ).rep == DMP([ZZ(1), ZZ(5)], ZZ)
    assert Poly.from_expr(y + 5, domain=ZZ).rep == DMP([ZZ(1), ZZ(5)], ZZ)

    assert Poly.from_expr(x + 5, x, domain=ZZ).rep == DMP([ZZ(1), ZZ(5)], ZZ)
    assert Poly.from_expr(y + 5, y, domain=ZZ).rep == DMP([ZZ(1), ZZ(5)], ZZ)

    assert Poly.from_expr(x + 5, x, y, domain=ZZ).rep == DMP([[ZZ(1)], [ZZ(5)]], ZZ)
    assert Poly.from_expr(y + 5, x, y, domain=ZZ).rep == DMP([[ZZ(1), ZZ(5)]], ZZ)


def test_Poly_rootof_extension():
    r1 = rootof(x**3 + x + 3, 0)
    r2 = rootof(x**3 + x + 3, 1)
    K1 = QQ.algebraic_field(r1)
    K2 = QQ.algebraic_field(r2)
    assert Poly(r1, y) == Poly(r1, y, domain=EX)
    assert Poly(r2, y) == Poly(r2, y, domain=EX)
    assert Poly(r1, y, extension=True) == Poly(r1, y, domain=K1)
    assert Poly(r2, y, extension=True) == Poly(r2, y, domain=K2)


@tooslow
def test_Poly_rootof_extension_primitive_element():
    r1 = rootof(x**3 + x + 3, 0)
    r2 = rootof(x**3 + x + 3, 1)
    K12 = QQ.algebraic_field(r1 + r2)
    assert Poly(r1*y + r2, y, extension=True) == Poly(r1*y + r2, y, domain=K12)


@XFAIL
def test_Poly_rootof_same_symbol_issue_26808():
    # XXX: This fails because r1 contains x.
    r1 = rootof(x**3 + x + 3, 0)
    K1 = QQ.algebraic_field(r1)
    assert Poly(r1, x) == Poly(r1, x, domain=EX)
    assert Poly(r1, x, extension=True) == Poly(r1, x, domain=K1)


def test_Poly_rootof_extension_to_sympy():
    # Verify that when primitive elements and RootOf are used, the expression
    # is not exploded on the way back to sympy.
    r1 = rootof(y**3 + y**2 - 1, 0)
    r2 = rootof(z**5 + z**2 - 1, 0)
    p = -x**5 + x**2 + x*r1 - r2 + 3*r1**2
    assert p.as_poly(x, extension=True).as_expr() == p


def test_poly_from_domain_element():
    dom = ZZ[x]
    assert Poly(dom(x+1), y, domain=dom).rep == DMP([dom(x+1)], dom)
    dom = dom.get_field()
    assert Poly(dom(x+1), y, domain=dom).rep == DMP([dom(x+1)], dom)

    dom = QQ[x]
    assert Poly(dom(x+1), y, domain=dom).rep == DMP([dom(x+1)], dom)
    dom = dom.get_field()
    assert Poly(dom(x+1), y, domain=dom).rep == DMP([dom(x+1)], dom)

    dom = ZZ.old_poly_ring(x)
    assert Poly(dom([ZZ(1), ZZ(1)]), y, domain=dom).rep == DMP([dom([ZZ(1), ZZ(1)])], dom)
    dom = dom.get_field()
    assert Poly(dom([ZZ(1), ZZ(1)]), y, domain=dom).rep == DMP([dom([ZZ(1), ZZ(1)])], dom)

    dom = QQ.old_poly_ring(x)
    assert Poly(dom([QQ(1), QQ(1)]), y, domain=dom).rep == DMP([dom([QQ(1), QQ(1)])], dom)
    dom = dom.get_field()
    assert Poly(dom([QQ(1), QQ(1)]), y, domain=dom).rep == DMP([dom([QQ(1), QQ(1)])], dom)

    dom = QQ.algebraic_field(I)
    assert Poly(dom([1, 1]), x, domain=dom).rep == DMP([dom([1, 1])], dom)


def test_Poly__new__():
    raises(GeneratorsError, lambda: Poly(x + 1, x, x))

    raises(GeneratorsError, lambda: Poly(x + y, x, y, domain=ZZ[x]))
    raises(GeneratorsError, lambda: Poly(x + y, x, y, domain=ZZ[y]))

    raises(OptionError, lambda: Poly(x, x, symmetric=True))
    raises(OptionError, lambda: Poly(x + 2, x, modulus=3, domain=QQ))

    raises(OptionError, lambda: Poly(x + 2, x, domain=ZZ, gaussian=True))
    raises(OptionError, lambda: Poly(x + 2, x, modulus=3, gaussian=True))

    raises(OptionError, lambda: Poly(x + 2, x, domain=ZZ, extension=[sqrt(3)]))
    raises(OptionError, lambda: Poly(x + 2, x, modulus=3, extension=[sqrt(3)]))

    raises(OptionError, lambda: Poly(x + 2, x, domain=ZZ, extension=True))
    raises(OptionError, lambda: Poly(x + 2, x, modulus=3, extension=True))

    raises(OptionError, lambda: Poly(x + 2, x, domain=ZZ, greedy=True))
    raises(OptionError, lambda: Poly(x + 2, x, domain=QQ, field=True))

    raises(OptionError, lambda: Poly(x + 2, x, domain=ZZ, greedy=False))
    raises(OptionError, lambda: Poly(x + 2, x, domain=QQ, field=False))

    raises(NotImplementedError, lambda: Poly(x + 1, x, modulus=3, order='grlex'))
    raises(NotImplementedError, lambda: Poly(x + 1, x, order='grlex'))

    raises(GeneratorsNeeded, lambda: Poly({1: 2, 0: 1}))
    raises(GeneratorsNeeded, lambda: Poly([2, 1]))
    raises(GeneratorsNeeded, lambda: Poly((2, 1)))

    raises(GeneratorsNeeded, lambda: Poly(1))

    assert Poly('x-x') == Poly(0, x)

    f = a*x**2 + b*x + c

    assert Poly({2: a, 1: b, 0: c}, x) == f
    assert Poly(iter([a, b, c]), x) == f
    assert Poly([a, b, c], x) == f
    assert Poly((a, b, c), x) == f

    f = Poly({}, x, y, z)

    assert f.gens == (x, y, z) and f.as_expr() == 0

    assert Poly(Poly(a*x + b*y, x, y), x) == Poly(a*x + b*y, x)

    assert Poly(3*x**2 + 2*x + 1, domain='ZZ').all_coeffs() == [3, 2, 1]
    assert Poly(3*x**2 + 2*x + 1, domain='QQ').all_coeffs() == [3, 2, 1]
    assert Poly(3*x**2 + 2*x + 1, domain='RR').all_coeffs() == [3.0, 2.0, 1.0]

    raises(CoercionFailed, lambda: Poly(3*x**2/5 + x*Rational(2, 5) + 1, domain='ZZ'))
    assert Poly(
        3*x**2/5 + x*Rational(2, 5) + 1, domain='QQ').all_coeffs() == [Rational(3, 5), Rational(2, 5), 1]
    assert _epsilon_eq(
        Poly(3*x**2/5 + x*Rational(2, 5) + 1, domain='RR').all_coeffs(), [0.6, 0.4, 1.0])

    assert Poly(3.0*x**2 + 2.0*x + 1, domain='ZZ').all_coeffs() == [3, 2, 1]
    assert Poly(3.0*x**2 + 2.0*x + 1, domain='QQ').all_coeffs() == [3, 2, 1]
    assert Poly(
        3.0*x**2 + 2.0*x + 1, domain='RR').all_coeffs() == [3.0, 2.0, 1.0]

    raises(CoercionFailed, lambda: Poly(3.1*x**2 + 2.1*x + 1, domain='ZZ'))
    assert Poly(3.1*x**2 + 2.1*x + 1, domain='QQ').all_coeffs() == [Rational(31, 10), Rational(21, 10), 1]
    assert Poly(3.1*x**2 + 2.1*x + 1, domain='RR').all_coeffs() == [3.1, 2.1, 1.0]

    assert Poly({(2, 1): 1, (1, 2): 2, (1, 1): 3}, x, y) == \
        Poly(x**2*y + 2*x*y**2 + 3*x*y, x, y)

    assert Poly(x**2 + 1, extension=I).get_domain() == QQ.algebraic_field(I)

    f = 3*x**5 - x**4 + x**3 - x** 2 + 65538

    assert Poly(f, x, modulus=65537, symmetric=True) == \
        Poly(3*x**5 - x**4 + x**3 - x** 2 + 1, x, modulus=65537,
             symmetric=True)
    assert Poly(f, x, modulus=65537, symmetric=False) == \
        Poly(3*x**5 + 65536*x**4 + x**3 + 65536*x** 2 + 1, x,
             modulus=65537, symmetric=False)

    N = 10**100
    assert Poly(-1, x, modulus=N, symmetric=False).as_expr() == N - 1

    assert isinstance(Poly(x**2 + x + 1.0).get_domain(), RealField)
    assert isinstance(Poly(x**2 + x + I + 1.0).get_domain(), ComplexField)


def test_Poly__args():
    assert Poly(x**2 + 1).args == (x**2 + 1, x)


def test_Poly__gens():
    assert Poly((x - p)*(x - q), x).gens == (x,)
    assert Poly((x - p)*(x - q), p).gens == (p,)
    assert Poly((x - p)*(x - q), q).gens == (q,)

    assert Poly((x - p)*(x - q), x, p).gens == (x, p)
    assert Poly((x - p)*(x - q), x, q).gens == (x, q)

    assert Poly((x - p)*(x - q), x, p, q).gens == (x, p, q)
    assert Poly((x - p)*(x - q), p, x, q).gens == (p, x, q)
    assert Poly((x - p)*(x - q), p, q, x).gens == (p, q, x)

    assert Poly((x - p)*(x - q)).gens == (x, p, q)

    assert Poly((x - p)*(x - q), sort='x > p > q').gens == (x, p, q)
    assert Poly((x - p)*(x - q), sort='p > x > q').gens == (p, x, q)
    assert Poly((x - p)*(x - q), sort='p > q > x').gens == (p, q, x)

    assert Poly((x - p)*(x - q), x, p, q, sort='p > q > x').gens == (x, p, q)

    assert Poly((x - p)*(x - q), wrt='x').gens == (x, p, q)
    assert Poly((x - p)*(x - q), wrt='p').gens == (p, x, q)
    assert Poly((x - p)*(x - q), wrt='q').gens == (q, x, p)

    assert Poly((x - p)*(x - q), wrt=x).gens == (x, p, q)
    assert Poly((x - p)*(x - q), wrt=p).gens == (p, x, q)
    assert Poly((x - p)*(x - q), wrt=q).gens == (q, x, p)

    assert Poly((x - p)*(x - q), x, p, q, wrt='p').gens == (x, p, q)

    assert Poly((x - p)*(x - q), wrt='p', sort='q > x').gens == (p, q, x)
    assert Poly((x - p)*(x - q), wrt='q', sort='p > x').gens == (q, p, x)


def test_Poly_zero():
    assert Poly(x).zero == Poly(0, x, domain=ZZ)
    assert Poly(x/2).zero == Poly(0, x, domain=QQ)


def test_Poly_one():
    assert Poly(x).one == Poly(1, x, domain=ZZ)
    assert Poly(x/2).one == Poly(1, x, domain=QQ)


def test_Poly__unify():
    raises(UnificationFailed, lambda: Poly(x)._unify(y))

    F3 = FF(3)

    assert Poly(x, x, modulus=3)._unify(Poly(y, y, modulus=3))[2:] == (
        DMP([[F3(1)], []], F3), DMP([[F3(1), F3(0)]], F3))
    raises(UnificationFailed, lambda: Poly(x, x, modulus=3)._unify(Poly(y, y, modulus=5)))

    raises(UnificationFailed, lambda: Poly(y, x, y)._unify(Poly(x, x, modulus=3)))
    raises(UnificationFailed, lambda: Poly(x, x, modulus=3)._unify(Poly(y, x, y)))

    assert Poly(x + 1, x)._unify(Poly(x + 2, x))[2:] ==\
        (DMP([ZZ(1), ZZ(1)], ZZ), DMP([ZZ(1), ZZ(2)], ZZ))
    assert Poly(x + 1, x, domain='QQ')._unify(Poly(x + 2, x))[2:] ==\
        (DMP([QQ(1), QQ(1)], QQ), DMP([QQ(1), QQ(2)], QQ))
    assert Poly(x + 1, x)._unify(Poly(x + 2, x, domain='QQ'))[2:] ==\
        (DMP([QQ(1), QQ(1)], QQ), DMP([QQ(1), QQ(2)], QQ))

    assert Poly(x + 1, x)._unify(Poly(x + 2, x, y))[2:] ==\
        (DMP([[ZZ(1)], [ZZ(1)]], ZZ), DMP([[ZZ(1)], [ZZ(2)]], ZZ))
    assert Poly(x + 1, x, domain='QQ')._unify(Poly(x + 2, x, y))[2:] ==\
        (DMP([[QQ(1)], [QQ(1)]], QQ), DMP([[QQ(1)], [QQ(2)]], QQ))
    assert Poly(x + 1, x)._unify(Poly(x + 2, x, y, domain='QQ'))[2:] ==\
        (DMP([[QQ(1)], [QQ(1)]], QQ), DMP([[QQ(1)], [QQ(2)]], QQ))

    assert Poly(x + 1, x, y)._unify(Poly(x + 2, x))[2:] ==\
        (DMP([[ZZ(1)], [ZZ(1)]], ZZ), DMP([[ZZ(1)], [ZZ(2)]], ZZ))
    assert Poly(x + 1, x, y, domain='QQ')._unify(Poly(x + 2, x))[2:] ==\
        (DMP([[QQ(1)], [QQ(1)]], QQ), DMP([[QQ(1)], [QQ(2)]], QQ))
    assert Poly(x + 1, x, y)._unify(Poly(x + 2, x, domain='QQ'))[2:] ==\
        (DMP([[QQ(1)], [QQ(1)]], QQ), DMP([[QQ(1)], [QQ(2)]], QQ))

    assert Poly(x + 1, x, y)._unify(Poly(x + 2, x, y))[2:] ==\
        (DMP([[ZZ(1)], [ZZ(1)]], ZZ), DMP([[ZZ(1)], [ZZ(2)]], ZZ))
    assert Poly(x + 1, x, y, domain='QQ')._unify(Poly(x + 2, x, y))[2:] ==\
        (DMP([[QQ(1)], [QQ(1)]], QQ), DMP([[QQ(1)], [QQ(2)]], QQ))
    assert Poly(x + 1, x, y)._unify(Poly(x + 2, x, y, domain='QQ'))[2:] ==\
        (DMP([[QQ(1)], [QQ(1)]], QQ), DMP([[QQ(1)], [QQ(2)]], QQ))

    assert Poly(x + 1, x)._unify(Poly(x + 2, y, x))[2:] ==\
        (DMP([[ZZ(1), ZZ(1)]], ZZ), DMP([[ZZ(1), ZZ(2)]], ZZ))
    assert Poly(x + 1, x, domain='QQ')._unify(Poly(x + 2, y, x))[2:] ==\
        (DMP([[QQ(1), QQ(1)]], QQ), DMP([[QQ(1), QQ(2)]], QQ))
    assert Poly(x + 1, x)._unify(Poly(x + 2, y, x, domain='QQ'))[2:] ==\
        (DMP([[QQ(1), QQ(1)]], QQ), DMP([[QQ(1), QQ(2)]], QQ))

    assert Poly(x + 1, y, x)._unify(Poly(x + 2, x))[2:] ==\
        (DMP([[ZZ(1), ZZ(1)]], ZZ), DMP([[ZZ(1), ZZ(2)]], ZZ))
    assert Poly(x + 1, y, x, domain='QQ')._unify(Poly(x + 2, x))[2:] ==\
        (DMP([[QQ(1), QQ(1)]], QQ), DMP([[QQ(1), QQ(2)]], QQ))
    assert Poly(x + 1, y, x)._unify(Poly(x + 2, x, domain='QQ'))[2:] ==\
        (DMP([[QQ(1), QQ(1)]], QQ), DMP([[QQ(1), QQ(2)]], QQ))

    assert Poly(x + 1, x, y)._unify(Poly(x + 2, y, x))[2:] ==\
        (DMP([[ZZ(1)], [ZZ(1)]], ZZ), DMP([[ZZ(1)], [ZZ(2)]], ZZ))
    assert Poly(x + 1, x, y, domain='QQ')._unify(Poly(x + 2, y, x))[2:] ==\
        (DMP([[QQ(1)], [QQ(1)]], QQ), DMP([[QQ(1)], [QQ(2)]], QQ))
    assert Poly(x + 1, x, y)._unify(Poly(x + 2, y, x, domain='QQ'))[2:] ==\
        (DMP([[QQ(1)], [QQ(1)]], QQ), DMP([[QQ(1)], [QQ(2)]], QQ))

    assert Poly(x + 1, y, x)._unify(Poly(x + 2, x, y))[2:] ==\
        (DMP([[ZZ(1), ZZ(1)]], ZZ), DMP([[ZZ(1), ZZ(2)]], ZZ))
    assert Poly(x + 1, y, x, domain='QQ')._unify(Poly(x + 2, x, y))[2:] ==\
        (DMP([[QQ(1), QQ(1)]], QQ), DMP([[QQ(1), QQ(2)]], QQ))
    assert Poly(x + 1, y, x)._unify(Poly(x + 2, x, y, domain='QQ'))[2:] ==\
        (DMP([[QQ(1), QQ(1)]], QQ), DMP([[QQ(1), QQ(2)]], QQ))

    assert Poly(x**2 + I, x, domain=ZZ_I).unify(Poly(x**2 + sqrt(2), x, extension=True)) == \
            (Poly(x**2 + I, x, domain='QQ<sqrt(2) + I>'), Poly(x**2 + sqrt(2), x, domain='QQ<sqrt(2) + I>'))

    F, A, B = field("a,b", ZZ)

    assert Poly(a*x, x, domain='ZZ[a]')._unify(Poly(a*b*x, x, domain='ZZ(a,b)'))[2:] == \
        (DMP([A, F(0)], F.to_domain()), DMP([A*B, F(0)], F.to_domain()))

    assert Poly(a*x, x, domain='ZZ(a)')._unify(Poly(a*b*x, x, domain='ZZ(a,b)'))[2:] == \
        (DMP([A, F(0)], F.to_domain()), DMP([A*B, F(0)], F.to_domain()))

    raises(CoercionFailed, lambda: Poly(Poly(x**2 + x**2*z, y, field=True), domain='ZZ(x)'))

    f = Poly(t**2 + t/3 + x, t, domain='QQ(x)')
    g = Poly(t**2 + t/3 + x, t, domain='QQ[x]')

    assert f._unify(g)[2:] == (f.rep, f.rep)


def test_Poly_free_symbols():
    assert Poly(x**2 + 1).free_symbols == {x}
    assert Poly(x**2 + y*z).free_symbols == {x, y, z}
    assert Poly(x**2 + y*z, x).free_symbols == {x, y, z}
    assert Poly(x**2 + sin(y*z)).free_symbols == {x, y, z}
    assert Poly(x**2 + sin(y*z), x).free_symbols == {x, y, z}
    assert Poly(x**2 + sin(y*z), x, domain=EX).free_symbols == {x, y, z}
    assert Poly(1 + x + x**2, x, y, z).free_symbols == {x}
    assert Poly(x + sin(y), z).free_symbols == {x, y}


def test_PurePoly_free_symbols():
    assert PurePoly(x**2 + 1).free_symbols == set()
    assert PurePoly(x**2 + y*z).free_symbols == set()
    assert PurePoly(x**2 + y*z, x).free_symbols == {y, z}
    assert PurePoly(x**2 + sin(y*z)).free_symbols == set()
    assert PurePoly(x**2 + sin(y*z), x).free_symbols == {y, z}
    assert PurePoly(x**2 + sin(y*z), x, domain=EX).free_symbols == {y, z}


def test_Poly__eq__():
    assert (Poly(x, x) == Poly(x, x)) is True
    assert (Poly(x, x, domain=QQ) == Poly(x, x)) is False
    assert (Poly(x, x) == Poly(x, x, domain=QQ)) is False

    assert (Poly(x, x, domain=ZZ[a]) == Poly(x, x)) is False
    assert (Poly(x, x) == Poly(x, x, domain=ZZ[a])) is False

    assert (Poly(x*y, x, y) == Poly(x, x)) is False

    assert (Poly(x, x, y) == Poly(x, x)) is False
    assert (Poly(x, x) == Poly(x, x, y)) is False

    assert (Poly(x**2 + 1, x) == Poly(y**2 + 1, y)) is False
    assert (Poly(y**2 + 1, y) == Poly(x**2 + 1, x)) is False

    f = Poly(x, x, domain=ZZ)
    g = Poly(x, x, domain=QQ)

    assert f.eq(g) is False
    assert f.ne(g) is True

    assert f.eq(g, strict=True) is False
    assert f.ne(g, strict=True) is True

    t0 = Symbol('t0')

    f =  Poly((t0/2 + x**2)*t**2 - x**2*t, t, domain='QQ[x,t0]')
    g =  Poly((t0/2 + x**2)*t**2 - x**2*t, t, domain='ZZ(x,t0)')

    assert (f == g) is False


def test_PurePoly__eq__():
    assert (PurePoly(x, x) == PurePoly(x, x)) is True
    assert (PurePoly(x, x, domain=QQ) == PurePoly(x, x)) is True
    assert (PurePoly(x, x) == PurePoly(x, x, domain=QQ)) is True

    assert (PurePoly(x, x, domain=ZZ[a]) == PurePoly(x, x)) is True
    assert (PurePoly(x, x) == PurePoly(x, x, domain=ZZ[a])) is True

    assert (PurePoly(x*y, x, y) == PurePoly(x, x)) is False

    assert (PurePoly(x, x, y) == PurePoly(x, x)) is False
    assert (PurePoly(x, x) == PurePoly(x, x, y)) is False

    assert (PurePoly(x**2 + 1, x) == PurePoly(y**2 + 1, y)) is True
    assert (PurePoly(y**2 + 1, y) == PurePoly(x**2 + 1, x)) is True

    f = PurePoly(x, x, domain=ZZ)
    g = PurePoly(x, x, domain=QQ)

    assert f.eq(g) is True
    assert f.ne(g) is False

    assert f.eq(g, strict=True) is False
    assert f.ne(g, strict=True) is True

    f = PurePoly(x, x, domain=ZZ)
    g = PurePoly(y, y, domain=QQ)

    assert f.eq(g) is True
    assert f.ne(g) is False

    assert f.eq(g, strict=True) is False
    assert f.ne(g, strict=True) is True


def test_PurePoly_Poly():
    assert isinstance(PurePoly(Poly(x**2 + 1)), PurePoly) is True
    assert isinstance(Poly(PurePoly(x**2 + 1)), Poly) is True


def test_Poly_get_domain():
    assert Poly(2*x).get_domain() == ZZ

    assert Poly(2*x, domain='ZZ').get_domain() == ZZ
    assert Poly(2*x, domain='QQ').get_domain() == QQ

    assert Poly(x/2).get_domain() == QQ

    raises(CoercionFailed, lambda: Poly(x/2, domain='ZZ'))
    assert Poly(x/2, domain='QQ').get_domain() == QQ

    assert isinstance(Poly(0.2*x).get_domain(), RealField)


def test_Poly_set_domain():
    assert Poly(2*x + 1).set_domain(ZZ) == Poly(2*x + 1)
    assert Poly(2*x + 1).set_domain('ZZ') == Poly(2*x + 1)

    assert Poly(2*x + 1).set_domain(QQ) == Poly(2*x + 1, domain='QQ')
    assert Poly(2*x + 1).set_domain('QQ') == Poly(2*x + 1, domain='QQ')

    assert Poly(Rational(2, 10)*x + Rational(1, 10)).set_domain('RR') == Poly(0.2*x + 0.1)
    assert Poly(0.2*x + 0.1).set_domain('QQ') == Poly(Rational(2, 10)*x + Rational(1, 10))

    raises(CoercionFailed, lambda: Poly(x/2 + 1).set_domain(ZZ))
    raises(CoercionFailed, lambda: Poly(x + 1, modulus=2).set_domain(QQ))

    raises(GeneratorsError, lambda: Poly(x*y, x, y).set_domain(ZZ[y]))


def test_Poly_get_modulus():
    assert Poly(x**2 + 1, modulus=2).get_modulus() == 2
    raises(PolynomialError, lambda: Poly(x**2 + 1).get_modulus())


def test_Poly_set_modulus():
    assert Poly(
        x**2 + 1, modulus=2).set_modulus(7) == Poly(x**2 + 1, modulus=7)
    assert Poly(
        x**2 + 5, modulus=7).set_modulus(2) == Poly(x**2 + 1, modulus=2)

    assert Poly(x**2 + 1).set_modulus(2) == Poly(x**2 + 1, modulus=2)

    raises(CoercionFailed, lambda: Poly(x/2 + 1).set_modulus(2))


def test_Poly_add_ground():
    assert Poly(x + 1).add_ground(2) == Poly(x + 3)


def test_Poly_sub_ground():
    assert Poly(x + 1).sub_ground(2) == Poly(x - 1)


def test_Poly_mul_ground():
    assert Poly(x + 1).mul_ground(2) == Poly(2*x + 2)


def test_Poly_quo_ground():
    assert Poly(2*x + 4).quo_ground(2) == Poly(x + 2)
    assert Poly(2*x + 3).quo_ground(2) == Poly(x + 1)


def test_Poly_exquo_ground():
    assert Poly(2*x + 4).exquo_ground(2) == Poly(x + 2)
    raises(ExactQuotientFailed, lambda: Poly(2*x + 3).exquo_ground(2))


def test_Poly_abs():
    assert Poly(-x + 1, x).abs() == abs(Poly(-x + 1, x)) == Poly(x + 1, x)


def test_Poly_neg():
    assert Poly(-x + 1, x).neg() == -Poly(-x + 1, x) == Poly(x - 1, x)


def test_Poly_add():
    assert Poly(0, x).add(Poly(0, x)) == Poly(0, x)
    assert Poly(0, x) + Poly(0, x) == Poly(0, x)

    assert Poly(1, x).add(Poly(0, x)) == Poly(1, x)
    assert Poly(1, x, y) + Poly(0, x) == Poly(1, x, y)
    assert Poly(0, x).add(Poly(1, x, y)) == Poly(1, x, y)
    assert Poly(0, x, y) + Poly(1, x, y) == Poly(1, x, y)

    assert Poly(1, x) + x == Poly(x + 1, x)
    with warns_deprecated_sympy():
        Poly(1, x) + sin(x)

    assert Poly(x, x) + 1 == Poly(x + 1, x)
    assert 1 + Poly(x, x) == Poly(x + 1, x)


def test_Poly_sub():
    assert Poly(0, x).sub(Poly(0, x)) == Poly(0, x)
    assert Poly(0, x) - Poly(0, x) == Poly(0, x)

    assert Poly(1, x).sub(Poly(0, x)) == Poly(1, x)
    assert Poly(1, x, y) - Poly(0, x) == Poly(1, x, y)
    assert Poly(0, x).sub(Poly(1, x, y)) == Poly(-1, x, y)
    assert Poly(0, x, y) - Poly(1, x, y) == Poly(-1, x, y)

    assert Poly(1, x) - x == Poly(1 - x, x)
    with warns_deprecated_sympy():
        Poly(1, x) - sin(x)

    assert Poly(x, x) - 1 == Poly(x - 1, x)
    assert 1 - Poly(x, x) == Poly(1 - x, x)


def test_Poly_mul():
    assert Poly(0, x).mul(Poly(0, x)) == Poly(0, x)
    assert Poly(0, x) * Poly(0, x) == Poly(0, x)

    assert Poly(2, x).mul(Poly(4, x)) == Poly(8, x)
    assert Poly(2, x, y) * Poly(4, x) == Poly(8, x, y)
    assert Poly(4, x).mul(Poly(2, x, y)) == Poly(8, x, y)
    assert Poly(4, x, y) * Poly(2, x, y) == Poly(8, x, y)

    assert Poly(1, x) * x == Poly(x, x)
    with warns_deprecated_sympy():
        Poly(1, x) * sin(x)

    assert Poly(x, x) * 2 == Poly(2*x, x)
    assert 2 * Poly(x, x) == Poly(2*x, x)

def test_issue_13079():
    assert Poly(x)*x == Poly(x**2, x, domain='ZZ')
    assert x*Poly(x) == Poly(x**2, x, domain='ZZ')
    assert -2*Poly(x) == Poly(-2*x, x, domain='ZZ')
    assert S(-2)*Poly(x) == Poly(-2*x, x, domain='ZZ')
    assert Poly(x)*S(-2) == Poly(-2*x, x, domain='ZZ')

def test_Poly_sqr():
    assert Poly(x*y, x, y).sqr() == Poly(x**2*y**2, x, y)


def test_Poly_pow():
    assert Poly(x, x).pow(10) == Poly(x**10, x)
    assert Poly(x, x).pow(Integer(10)) == Poly(x**10, x)

    assert Poly(2*y, x, y).pow(4) == Poly(16*y**4, x, y)
    assert Poly(2*y, x, y).pow(Integer(4)) == Poly(16*y**4, x, y)

    assert Poly(7*x*y, x, y)**3 == Poly(343*x**3*y**3, x, y)

    raises(TypeError, lambda: Poly(x*y + 1, x, y)**(-1))
    raises(TypeError, lambda: Poly(x*y + 1, x, y)**x)


def test_Poly_divmod():
    f, g = Poly(x**2), Poly(x)
    q, r = g, Poly(0, x)

    assert divmod(f, g) == (q, r)
    assert f // g == q
    assert f % g == r

    assert divmod(f, x) == (q, r)
    assert f // x == q
    assert f % x == r

    q, r = Poly(0, x), Poly(2, x)

    assert divmod(2, g) == (q, r)
    assert 2 // g == q
    assert 2 % g == r

    assert Poly(x)/Poly(x) == 1
    assert Poly(x**2)/Poly(x) == x
    assert Poly(x)/Poly(x**2) == 1/x


def test_Poly_eq_ne():
    assert (Poly(x + y, x, y) == Poly(x + y, x, y)) is True
    assert (Poly(x + y, x) == Poly(x + y, x, y)) is False
    assert (Poly(x + y, x, y) == Poly(x + y, x)) is False
    assert (Poly(x + y, x) == Poly(x + y, x)) is True
    assert (Poly(x + y, y) == Poly(x + y, y)) is True

    assert (Poly(x + y, x, y) == x + y) is True
    assert (Poly(x + y, x) == x + y) is True
    assert (Poly(x + y, x, y) == x + y) is True
    assert (Poly(x + y, x) == x + y) is True
    assert (Poly(x + y, y) == x + y) is True

    assert (Poly(x + y, x, y) != Poly(x + y, x, y)) is False
    assert (Poly(x + y, x) != Poly(x + y, x, y)) is True
    assert (Poly(x + y, x, y) != Poly(x + y, x)) is True
    assert (Poly(x + y, x) != Poly(x + y, x)) is False
    assert (Poly(x + y, y) != Poly(x + y, y)) is False

    assert (Poly(x + y, x, y) != x + y) is False
    assert (Poly(x + y, x) != x + y) is False
    assert (Poly(x + y, x, y) != x + y) is False
    assert (Poly(x + y, x) != x + y) is False
    assert (Poly(x + y, y) != x + y) is False

    assert (Poly(x, x) == sin(x)) is False
    assert (Poly(x, x) != sin(x)) is True


def test_Poly_nonzero():
    assert not bool(Poly(0, x)) is True
    assert not bool(Poly(1, x)) is False


def test_Poly_properties():
    assert Poly(0, x).is_zero is True
    assert Poly(1, x).is_zero is False

    assert Poly(1, x).is_one is True
    assert Poly(2, x).is_one is False

    assert Poly(x - 1, x).is_sqf is True
    assert Poly((x - 1)**2, x).is_sqf is False

    assert Poly(x - 1, x).is_monic is True
    assert Poly(2*x - 1, x).is_monic is False

    assert Poly(3*x + 2, x).is_primitive is True
    assert Poly(4*x + 2, x).is_primitive is False

    assert Poly(1, x).is_ground is True
    assert Poly(x, x).is_ground is False

    assert Poly(x + y + z + 1).is_linear is True
    assert Poly(x*y*z + 1).is_linear is False

    assert Poly(x*y + z + 1).is_quadratic is True
    assert Poly(x*y*z + 1).is_quadratic is False

    assert Poly(x*y).is_monomial is True
    assert Poly(x*y + 1).is_monomial is False

    assert Poly(x**2 + x*y).is_homogeneous is True
    assert Poly(x**3 + x*y).is_homogeneous is False

    assert Poly(x).is_univariate is True
    assert Poly(x*y).is_univariate is False

    assert Poly(x*y).is_multivariate is True
    assert Poly(x).is_multivariate is False

    assert Poly(
        x**16 + x**14 - x**10 + x**8 - x**6 + x**2 + 1).is_cyclotomic is False
    assert Poly(
        x**16 + x**14 - x**10 - x**8 - x**6 + x**2 + 1).is_cyclotomic is True


def test_Poly_is_irreducible():
    assert Poly(x**2 + x + 1).is_irreducible is True
    assert Poly(x**2 + 2*x + 1).is_irreducible is False

    assert Poly(7*x + 3, modulus=11).is_irreducible is True
    assert Poly(7*x**2 + 3*x + 1, modulus=11).is_irreducible is False


def test_Poly_subs():
    assert Poly(x + 1).subs(x, 0) == 1

    assert Poly(x + 1).subs(x, x) == Poly(x + 1)
    assert Poly(x + 1).subs(x, y) == Poly(y + 1)

    assert Poly(x*y, x).subs(y, x) == x**2
    assert Poly(x*y, x).subs(x, y) == y**2


def test_Poly_replace():
    assert Poly(x + 1).replace(x) == Poly(x + 1)
    assert Poly(x + 1).replace(y) == Poly(y + 1)

    raises(PolynomialError, lambda: Poly(x + y).replace(z))

    assert Poly(x + 1).replace(x, x) == Poly(x + 1)
    assert Poly(x + 1).replace(x, y) == Poly(y + 1)

    assert Poly(x + y).replace(x, x) == Poly(x + y)
    assert Poly(x + y).replace(x, z) == Poly(z + y, z, y)

    assert Poly(x + y).replace(y, y) == Poly(x + y)
    assert Poly(x + y).replace(y, z) == Poly(x + z, x, z)
    assert Poly(x + y).replace(z, t) == Poly(x + y)

    raises(PolynomialError, lambda: Poly(x + y).replace(x, y))

    assert Poly(x + y, x).replace(x, z) == Poly(z + y, z)
    assert Poly(x + y, y).replace(y, z) == Poly(x + z, z)

    raises(PolynomialError, lambda: Poly(x + y, x).replace(x, y))
    raises(PolynomialError, lambda: Poly(x + y, y).replace(y, x))


def test_Poly_reorder():
    raises(PolynomialError, lambda: Poly(x + y).reorder(x, z))

    assert Poly(x + y, x, y).reorder(x, y) == Poly(x + y, x, y)
    assert Poly(x + y, x, y).reorder(y, x) == Poly(x + y, y, x)

    assert Poly(x + y, y, x).reorder(x, y) == Poly(x + y, x, y)
    assert Poly(x + y, y, x).reorder(y, x) == Poly(x + y, y, x)

    assert Poly(x + y, x, y).reorder(wrt=x) == Poly(x + y, x, y)
    assert Poly(x + y, x, y).reorder(wrt=y) == Poly(x + y, y, x)


def test_Poly_ltrim():
    f = Poly(y**2 + y*z**2, x, y, z).ltrim(y)
    assert f.as_expr() == y**2 + y*z**2 and f.gens == (y, z)
    assert Poly(x*y - x, z, x, y).ltrim(1) == Poly(x*y - x, x, y)

    raises(PolynomialError, lambda: Poly(x*y**2 + y**2, x, y).ltrim(y))
    raises(PolynomialError, lambda: Poly(x*y - x, x, y).ltrim(-1))

def test_Poly_has_only_gens():
    assert Poly(x*y + 1, x, y, z).has_only_gens(x, y) is True
    assert Poly(x*y + z, x, y, z).has_only_gens(x, y) is False

    raises(GeneratorsError, lambda: Poly(x*y**2 + y**2, x, y).has_only_gens(t))


def test_Poly_to_ring():
    assert Poly(2*x + 1, domain='ZZ').to_ring() == Poly(2*x + 1, domain='ZZ')
    assert Poly(2*x + 1, domain='QQ').to_ring() == Poly(2*x + 1, domain='ZZ')

    raises(CoercionFailed, lambda: Poly(x/2 + 1).to_ring())
    raises(DomainError, lambda: Poly(2*x + 1, modulus=3).to_ring())


def test_Poly_to_field():
    assert Poly(2*x + 1, domain='ZZ').to_field() == Poly(2*x + 1, domain='QQ')
    assert Poly(2*x + 1, domain='QQ').to_field() == Poly(2*x + 1, domain='QQ')

    assert Poly(x/2 + 1, domain='QQ').to_field() == Poly(x/2 + 1, domain='QQ')
    assert Poly(2*x + 1, modulus=3).to_field() == Poly(2*x + 1, modulus=3)

    assert Poly(2.0*x + 1.0).to_field() == Poly(2.0*x + 1.0)


def test_Poly_to_exact():
    assert Poly(2*x).to_exact() == Poly(2*x)
    assert Poly(x/2).to_exact() == Poly(x/2)

    assert Poly(0.1*x).to_exact() == Poly(x/10)


def test_Poly_retract():
    f = Poly(x**2 + 1, x, domain=QQ[y])

    assert f.retract() == Poly(x**2 + 1, x, domain='ZZ')
    assert f.retract(field=True) == Poly(x**2 + 1, x, domain='QQ')

    assert Poly(0, x, y).retract() == Poly(0, x, y)


def test_Poly_slice():
    f = Poly(x**3 + 2*x**2 + 3*x + 4)

    assert f.slice(0, 0) == Poly(0, x)
    assert f.slice(0, 1) == Poly(4, x)
    assert f.slice(0, 2) == Poly(3*x + 4, x)
    assert f.slice(0, 3) == Poly(2*x**2 + 3*x + 4, x)
    assert f.slice(0, 4) == Poly(x**3 + 2*x**2 + 3*x + 4, x)

    assert f.slice(x, 0, 0) == Poly(0, x)
    assert f.slice(x, 0, 1) == Poly(4, x)
    assert f.slice(x, 0, 2) == Poly(3*x + 4, x)
    assert f.slice(x, 0, 3) == Poly(2*x**2 + 3*x + 4, x)
    assert f.slice(x, 0, 4) == Poly(x**3 + 2*x**2 + 3*x + 4, x)

    g = Poly(x**3 + 1)

    assert g.slice(0, 3) == Poly(1, x)


def test_Poly_coeffs():
    assert Poly(0, x).coeffs() == [0]
    assert Poly(1, x).coeffs() == [1]

    assert Poly(2*x + 1, x).coeffs() == [2, 1]

    assert Poly(7*x**2 + 2*x + 1, x).coeffs() == [7, 2, 1]
    assert Poly(7*x**4 + 2*x + 1, x).coeffs() == [7, 2, 1]

    assert Poly(x*y**7 + 2*x**2*y**3).coeffs('lex') == [2, 1]
    assert Poly(x*y**7 + 2*x**2*y**3).coeffs('grlex') == [1, 2]


def test_Poly_monoms():
    assert Poly(0, x).monoms() == [(0,)]
    assert Poly(1, x).monoms() == [(0,)]

    assert Poly(2*x + 1, x).monoms() == [(1,), (0,)]

    assert Poly(7*x**2 + 2*x + 1, x).monoms() == [(2,), (1,), (0,)]
    assert Poly(7*x**4 + 2*x + 1, x).monoms() == [(4,), (1,), (0,)]

    assert Poly(x*y**7 + 2*x**2*y**3).monoms('lex') == [(2, 3), (1, 7)]
    assert Poly(x*y**7 + 2*x**2*y**3).monoms('grlex') == [(1, 7), (2, 3)]


def test_Poly_terms():
    assert Poly(0, x).terms() == [((0,), 0)]
    assert Poly(1, x).terms() == [((0,), 1)]

    assert Poly(2*x + 1, x).terms() == [((1,), 2), ((0,), 1)]

    assert Poly(7*x**2 + 2*x + 1, x).terms() == [((2,), 7), ((1,), 2), ((0,), 1)]
    assert Poly(7*x**4 + 2*x + 1, x).terms() == [((4,), 7), ((1,), 2), ((0,), 1)]

    assert Poly(
        x*y**7 + 2*x**2*y**3).terms('lex') == [((2, 3), 2), ((1, 7), 1)]
    assert Poly(
        x*y**7 + 2*x**2*y**3).terms('grlex') == [((1, 7), 1), ((2, 3), 2)]


def test_Poly_all_coeffs():
    assert Poly(0, x).all_coeffs() == [0]
    assert Poly(1, x).all_coeffs() == [1]

    assert Poly(2*x + 1, x).all_coeffs() == [2, 1]

    assert Poly(7*x**2 + 2*x + 1, x).all_coeffs() == [7, 2, 1]
    assert Poly(7*x**4 + 2*x + 1, x).all_coeffs() == [7, 0, 0, 2, 1]


def test_Poly_all_monoms():
    assert Poly(0, x).all_monoms() == [(0,)]
    assert Poly(1, x).all_monoms() == [(0,)]

    assert Poly(2*x + 1, x).all_monoms() == [(1,), (0,)]

    assert Poly(7*x**2 + 2*x + 1, x).all_monoms() == [(2,), (1,), (0,)]
    assert Poly(7*x**4 + 2*x + 1, x).all_monoms() == [(4,), (3,), (2,), (1,), (0,)]


def test_Poly_all_terms():
    assert Poly(0, x).all_terms() == [((0,), 0)]
    assert Poly(1, x).all_terms() == [((0,), 1)]

    assert Poly(2*x + 1, x).all_terms() == [((1,), 2), ((0,), 1)]

    assert Poly(7*x**2 + 2*x + 1, x).all_terms() == \
        [((2,), 7), ((1,), 2), ((0,), 1)]
    assert Poly(7*x**4 + 2*x + 1, x).all_terms() == \
        [((4,), 7), ((3,), 0), ((2,), 0), ((1,), 2), ((0,), 1)]


def test_Poly_termwise():
    f = Poly(x**2 + 20*x + 400)
    g = Poly(x**2 + 2*x + 4)

    def func(monom, coeff):
        (k,) = monom
        return coeff//10**(2 - k)

    assert f.termwise(func) == g

    def func(monom, coeff):
        (k,) = monom
        return (k,), coeff//10**(2 - k)

    assert f.termwise(func) == g


def test_Poly_length():
    assert Poly(0, x).length() == 0
    assert Poly(1, x).length() == 1
    assert Poly(x, x).length() == 1

    assert Poly(x + 1, x).length() == 2
    assert Poly(x**2 + 1, x).length() == 2
    assert Poly(x**2 + x + 1, x).length() == 3


def test_Poly_as_dict():
    assert Poly(0, x).as_dict() == {}
    assert Poly(0, x, y, z).as_dict() == {}

    assert Poly(1, x).as_dict() == {(0,): 1}
    assert Poly(1, x, y, z).as_dict() == {(0, 0, 0): 1}

    assert Poly(x**2 + 3, x).as_dict() == {(2,): 1, (0,): 3}
    assert Poly(x**2 + 3, x, y, z).as_dict() == {(2, 0, 0): 1, (0, 0, 0): 3}

    assert Poly(3*x**2*y*z**3 + 4*x*y + 5*x*z).as_dict() == {(2, 1, 3): 3,
                (1, 1, 0): 4, (1, 0, 1): 5}


def test_Poly_as_expr():
    assert Poly(0, x).as_expr() == 0
    assert Poly(0, x, y, z).as_expr() == 0

    assert Poly(1, x).as_expr() == 1
    assert Poly(1, x, y, z).as_expr() == 1

    assert Poly(x**2 + 3, x).as_expr() == x**2 + 3
    assert Poly(x**2 + 3, x, y, z).as_expr() == x**2 + 3

    assert Poly(
        3*x**2*y*z**3 + 4*x*y + 5*x*z).as_expr() == 3*x**2*y*z**3 + 4*x*y + 5*x*z

    f = Poly(x**2 + 2*x*y**2 - y, x, y)

    assert f.as_expr() == -y + x**2 + 2*x*y**2

    assert f.as_expr({x: 5}) == 25 - y + 10*y**2
    assert f.as_expr({y: 6}) == -6 + 72*x + x**2

    assert f.as_expr({x: 5, y: 6}) == 379
    assert f.as_expr(5, 6) == 379

    raises(GeneratorsError, lambda: f.as_expr({z: 7}))


def test_Poly_lift():
    p = Poly(x**4 - I*x + 17*I, x, gaussian=True)
    assert p.lift() == Poly(x**8 + x**2 - 34*x + 289, x, domain='QQ')


def test_Poly_lift_multiple():

    r1 = rootof(y**3 + y**2 - 1, 0)
    r2 = rootof(z**5 + z**2 - 1, 0)
    p = Poly(r1*x + 3*r1**2 - r2 + x**2 - x**5, x, extension=True)

    assert p.lift() == Poly(
        -x**75 + 15*x**72 - 5*x**71 + 15*x**70 - 105*x**69 + 70*x**68 -
        220*x**67 + 560*x**66 - 635*x**65 + 1495*x**64 - 2735*x**63 +
        4415*x**62 - 7410*x**61 + 12741*x**60 - 22090*x**59 + 32125*x**58 -
        56281*x**57 + 88157*x**56 - 126842*x**55 + 214223*x**54 - 311802*x**53
        + 462667*x**52 - 700883*x**51 + 1006278*x**50 - 1480950*x**49 +
        2078055*x**48 - 3004675*x**47 + 4140410*x**46 - 5664222*x**45 +
        8029445*x**44 - 10528785*x**43 + 14309614*x**42 - 19032988*x**41 +
        24570573*x**40 - 32530459*x**39 + 41239581*x**38 - 52968051*x**37 +
        65891606*x**36 - 81997276*x**35 + 102530732*x**34 - 122009994*x**33 +
        150227996*x**32 - 176452478*x**31 + 206393768*x**30 - 245291426*x**29 +
        276598718*x**28 - 320005297*x**27 + 353649032*x**26
        - 393246309*x**25 + 434566186*x**24 - 460608964*x**23 + 508052079*x**22
        - 513976618*x**21 + 539374498*x**20 - 557851717*x**19 + 540788016*x**18
        - 564949060*x**17 + 520866566*x**16
        - 507861375*x**15 + 474999819*x**14 - 423619160*x**13 + 414540540*x**12
        - 322522367*x**11 + 311586511*x**10 - 238812299*x**9 + 184482053*x**8
        - 189265274*x**7 + 93619528*x**6 - 106852385*x**5 + 57294385*x**4 -
        26486666*x**3 + 42614683*x**2 - 1511583*x + 15975845, x, domain='QQ'
    )


def test_Poly_deflate():
    assert Poly(0, x).deflate() == ((1,), Poly(0, x))
    assert Poly(1, x).deflate() == ((1,), Poly(1, x))
    assert Poly(x, x).deflate() == ((1,), Poly(x, x))

    assert Poly(x**2, x).deflate() == ((2,), Poly(x, x))
    assert Poly(x**17, x).deflate() == ((17,), Poly(x, x))

    assert Poly(
        x**2*y*z**11 + x**4*z**11).deflate() == ((2, 1, 11), Poly(x*y*z + x**2*z))


def test_Poly_inject():
    f = Poly(x**2*y + x*y**3 + x*y + 1, x)

    assert f.inject() == Poly(x**2*y + x*y**3 + x*y + 1, x, y)
    assert f.inject(front=True) == Poly(y**3*x + y*x**2 + y*x + 1, y, x)


def test_Poly_eject():
    f = Poly(x**2*y + x*y**3 + x*y + 1, x, y)

    assert f.eject(x) == Poly(x*y**3 + (x**2 + x)*y + 1, y, domain='ZZ[x]')
    assert f.eject(y) == Poly(y*x**2 + (y**3 + y)*x + 1, x, domain='ZZ[y]')

    ex = x + y + z + t + w
    g = Poly(ex, x, y, z, t, w)

    assert g.eject(x) == Poly(ex, y, z, t, w, domain='ZZ[x]')
    assert g.eject(x, y) == Poly(ex, z, t, w, domain='ZZ[x, y]')
    assert g.eject(x, y, z) == Poly(ex, t, w, domain='ZZ[x, y, z]')
    assert g.eject(w) == Poly(ex, x, y, z, t, domain='ZZ[w]')
    assert g.eject(t, w) == Poly(ex, x, y, z, domain='ZZ[t, w]')
    assert g.eject(z, t, w) == Poly(ex, x, y, domain='ZZ[z, t, w]')

    raises(DomainError, lambda: Poly(x*y, x, y, domain=ZZ[z]).eject(y))
    raises(NotImplementedError, lambda: Poly(x*y, x, y, z).eject(y))


def test_Poly_exclude():
    assert Poly(x, x, y).exclude() == Poly(x, x)
    assert Poly(x*y, x, y).exclude() == Poly(x*y, x, y)
    assert Poly(1, x, y).exclude() == Poly(1, x, y)


def test_Poly__gen_to_level():
    assert Poly(1, x, y)._gen_to_level(-2) == 0
    assert Poly(1, x, y)._gen_to_level(-1) == 1
    assert Poly(1, x, y)._gen_to_level( 0) == 0
    assert Poly(1, x, y)._gen_to_level( 1) == 1

    raises(PolynomialError, lambda: Poly(1, x, y)._gen_to_level(-3))
    raises(PolynomialError, lambda: Poly(1, x, y)._gen_to_level( 2))

    assert Poly(1, x, y)._gen_to_level(x) == 0
    assert Poly(1, x, y)._gen_to_level(y) == 1

    assert Poly(1, x, y)._gen_to_level('x') == 0
    assert Poly(1, x, y)._gen_to_level('y') == 1

    raises(PolynomialError, lambda: Poly(1, x, y)._gen_to_level(z))
    raises(PolynomialError, lambda: Poly(1, x, y)._gen_to_level('z'))


def test_Poly_degree():
    assert Poly(0, x).degree() is -oo
    assert Poly(1, x).degree() == 0
    assert Poly(x, x).degree() == 1

    assert Poly(0, x).degree(gen=0) is -oo
    assert Poly(1, x).degree(gen=0) == 0
    assert Poly(x, x).degree(gen=0) == 1

    assert Poly(0, x).degree(gen=x) is -oo
    assert Poly(1, x).degree(gen=x) == 0
    assert Poly(x, x).degree(gen=x) == 1

    assert Poly(0, x).degree(gen='x') is -oo
    assert Poly(1, x).degree(gen='x') == 0
    assert Poly(x, x).degree(gen='x') == 1

    raises(PolynomialError, lambda: Poly(1, x).degree(gen=1))
    raises(PolynomialError, lambda: Poly(1, x).degree(gen=y))
    raises(PolynomialError, lambda: Poly(1, x).degree(gen='y'))

    assert Poly(1, x, y).degree() == 0
    assert Poly(2*y, x, y).degree() == 0
    assert Poly(x*y, x, y).degree() == 1

    assert Poly(1, x, y).degree(gen=x) == 0
    assert Poly(2*y, x, y).degree(gen=x) == 0
    assert Poly(x*y, x, y).degree(gen=x) == 1

    assert Poly(1, x, y).degree(gen=y) == 0
    assert Poly(2*y, x, y).degree(gen=y) == 1
    assert Poly(x*y, x, y).degree(gen=y) == 1

    assert degree(0, x) is -oo
    assert degree(1, x) == 0
    assert degree(x, x) == 1

    assert degree(x*y**2, x) == 1
    assert degree(x*y**2, y) == 2
    assert degree(x*y**2, z) == 0

    assert degree(pi) == 1

    raises(TypeError, lambda: degree(y**2 + x**3))
    raises(TypeError, lambda: degree(y**2 + x**3, 1))
    raises(PolynomialError, lambda: degree(x, 1.1))
    raises(PolynomialError, lambda: degree(x**2/(x**3 + 1), x))

    assert degree(Poly(0,x),z) is -oo
    assert degree(Poly(1,x),z) == 0
    assert degree(Poly(x**2+y**3,y)) == 3
    assert degree(Poly(y**2 + x**3, y, x), 1) == 3
    assert degree(Poly(y**2 + x**3, x), z) == 0
    assert degree(Poly(y**2 + x**3 + z**4, x), z) == 4

def test_Poly_degree_list():
    assert Poly(0, x).degree_list() == (-oo,)
    assert Poly(0, x, y).degree_list() == (-oo, -oo)
    assert Poly(0, x, y, z).degree_list() == (-oo, -oo, -oo)

    assert Poly(1, x).degree_list() == (0,)
    assert Poly(1, x, y).degree_list() == (0, 0)
    assert Poly(1, x, y, z).degree_list() == (0, 0, 0)

    assert Poly(x**2*y + x**3*z**2 + 1).degree_list() == (3, 1, 2)

    assert degree_list(1, x) == (0,)
    assert degree_list(x, x) == (1,)

    assert degree_list(x*y**2) == (1, 2)

    raises(ComputationFailed, lambda: degree_list(1))


def test_Poly_total_degree():
    assert Poly(x**2*y + x**3*z**2 + 1).total_degree() == 5
    assert Poly(x**2 + z**3).total_degree() == 3
    assert Poly(x*y*z + z**4).total_degree() == 4
    assert Poly(x**3 + x + 1).total_degree() == 3

    assert total_degree(x*y + z**3) == 3
    assert total_degree(x*y + z**3, x, y) == 2
    assert total_degree(1) == 0
    assert total_degree(Poly(y**2 + x**3 + z**4)) == 4
    assert total_degree(Poly(y**2 + x**3 + z**4, x)) == 3
    assert total_degree(Poly(y**2 + x**3 + z**4, x), z) == 4
    assert total_degree(Poly(x**9 + x*z*y + x**3*z**2 + z**7,x), z) == 7

def test_Poly_homogenize():
    assert Poly(x**2+y).homogenize(z) == Poly(x**2+y*z)
    assert Poly(x+y).homogenize(z) == Poly(x+y, x, y, z)
    assert Poly(x+y**2).homogenize(y) == Poly(x*y+y**2)


def test_Poly_homogeneous_order():
    assert Poly(0, x, y).homogeneous_order() is -oo
    assert Poly(1, x, y).homogeneous_order() == 0
    assert Poly(x, x, y).homogeneous_order() == 1
    assert Poly(x*y, x, y).homogeneous_order() == 2

    assert Poly(x + 1, x, y).homogeneous_order() is None
    assert Poly(x*y + x, x, y).homogeneous_order() is None

    assert Poly(x**5 + 2*x**3*y**2 + 9*x*y**4).homogeneous_order() == 5
    assert Poly(x**5 + 2*x**3*y**3 + 9*x*y**4).homogeneous_order() is None


def test_Poly_LC():
    assert Poly(0, x).LC() == 0
    assert Poly(1, x).LC() == 1
    assert Poly(2*x**2 + x, x).LC() == 2

    assert Poly(x*y**7 + 2*x**2*y**3).LC('lex') == 2
    assert Poly(x*y**7 + 2*x**2*y**3).LC('grlex') == 1

    assert LC(x*y**7 + 2*x**2*y**3, order='lex') == 2
    assert LC(x*y**7 + 2*x**2*y**3, order='grlex') == 1


def test_Poly_TC():
    assert Poly(0, x).TC() == 0
    assert Poly(1, x).TC() == 1
    assert Poly(2*x**2 + x, x).TC() == 0


def test_Poly_EC():
    assert Poly(0, x).EC() == 0
    assert Poly(1, x).EC() == 1
    assert Poly(2*x**2 + x, x).EC() == 1

    assert Poly(x*y**7 + 2*x**2*y**3).EC('lex') == 1
    assert Poly(x*y**7 + 2*x**2*y**3).EC('grlex') == 2


def test_Poly_coeff():
    assert Poly(0, x).coeff_monomial(1) == 0
    assert Poly(0, x).coeff_monomial(x) == 0

    assert Poly(1, x).coeff_monomial(1) == 1
    assert Poly(1, x).coeff_monomial(x) == 0

    assert Poly(x**8, x).coeff_monomial(1) == 0
    assert Poly(x**8, x).coeff_monomial(x**7) == 0
    assert Poly(x**8, x).coeff_monomial(x**8) == 1
    assert Poly(x**8, x).coeff_monomial(x**9) == 0

    assert Poly(3*x*y**2 + 1, x, y).coeff_monomial(1) == 1
    assert Poly(3*x*y**2 + 1, x, y).coeff_monomial(x*y**2) == 3

    p = Poly(24*x*y*exp(8) + 23*x, x, y)

    assert p.coeff_monomial(x) == 23
    assert p.coeff_monomial(y) == 0
    assert p.coeff_monomial(x*y) == 24*exp(8)

    assert p.as_expr().coeff(x) == 24*y*exp(8) + 23
    raises(NotImplementedError, lambda: p.coeff(x))

    raises(ValueError, lambda: Poly(x + 1).coeff_monomial(0))
    raises(ValueError, lambda: Poly(x + 1).coeff_monomial(3*x))
    raises(ValueError, lambda: Poly(x + 1).coeff_monomial(3*x*y))


def test_Poly_nth():
    assert Poly(0, x).nth(0) == 0
    assert Poly(0, x).nth(1) == 0

    assert Poly(1, x).nth(0) == 1
    assert Poly(1, x).nth(1) == 0

    assert Poly(x**8, x).nth(0) == 0
    assert Poly(x**8, x).nth(7) == 0
    assert Poly(x**8, x).nth(8) == 1
    assert Poly(x**8, x).nth(9) == 0

    assert Poly(3*x*y**2 + 1, x, y).nth(0, 0) == 1
    assert Poly(3*x*y**2 + 1, x, y).nth(1, 2) == 3

    raises(ValueError, lambda: Poly(x*y + 1, x, y).nth(1))


def test_Poly_LM():
    assert Poly(0, x).LM() == (0,)
    assert Poly(1, x).LM() == (0,)
    assert Poly(2*x**2 + x, x).LM() == (2,)

    assert Poly(x*y**7 + 2*x**2*y**3).LM('lex') == (2, 3)
    assert Poly(x*y**7 + 2*x**2*y**3).LM('grlex') == (1, 7)

    assert LM(x*y**7 + 2*x**2*y**3, order='lex') == x**2*y**3
    assert LM(x*y**7 + 2*x**2*y**3, order='grlex') == x*y**7


def test_Poly_LM_custom_order():
    f = Poly(x**2*y**3*z + x**2*y*z**3 + x*y*z + 1)
    rev_lex = lambda monom: tuple(reversed(monom))

    assert f.LM(order='lex') == (2, 3, 1)
    assert f.LM(order=rev_lex) == (2, 1, 3)


def test_Poly_EM():
    assert Poly(0, x).EM() == (0,)
    assert Poly(1, x).EM() == (0,)
    assert Poly(2*x**2 + x, x).EM() == (1,)

    assert Poly(x*y**7 + 2*x**2*y**3).EM('lex') == (1, 7)
    assert Poly(x*y**7 + 2*x**2*y**3).EM('grlex') == (2, 3)


def test_Poly_LT():
    assert Poly(0, x).LT() == ((0,), 0)
    assert Poly(1, x).LT() == ((0,), 1)
    assert Poly(2*x**2 + x, x).LT() == ((2,), 2)

    assert Poly(x*y**7 + 2*x**2*y**3).LT('lex') == ((2, 3), 2)
    assert Poly(x*y**7 + 2*x**2*y**3).LT('grlex') == ((1, 7), 1)

    assert LT(x*y**7 + 2*x**2*y**3, order='lex') == 2*x**2*y**3
    assert LT(x*y**7 + 2*x**2*y**3, order='grlex') == x*y**7


def test_Poly_ET():
    assert Poly(0, x).ET() == ((0,), 0)
    assert Poly(1, x).ET() == ((0,), 1)
    assert Poly(2*x**2 + x, x).ET() == ((1,), 1)

    assert Poly(x*y**7 + 2*x**2*y**3).ET('lex') == ((1, 7), 1)
    assert Poly(x*y**7 + 2*x**2*y**3).ET('grlex') == ((2, 3), 2)


def test_Poly_max_norm():
    assert Poly(-1, x).max_norm() == 1
    assert Poly( 0, x).max_norm() == 0
    assert Poly( 1, x).max_norm() == 1


def test_Poly_l1_norm():
    assert Poly(-1, x).l1_norm() == 1
    assert Poly( 0, x).l1_norm() == 0
    assert Poly( 1, x).l1_norm() == 1


def test_Poly_clear_denoms():
    coeff, poly = Poly(x + 2, x).clear_denoms()
    assert coeff == 1 and poly == Poly(
        x + 2, x, domain='ZZ') and poly.get_domain() == ZZ

    coeff, poly = Poly(x/2 + 1, x).clear_denoms()
    assert coeff == 2 and poly == Poly(
        x + 2, x, domain='QQ') and poly.get_domain() == QQ

    coeff, poly = Poly(2*x**2 + 3, modulus=5).clear_denoms()
    assert coeff == 1 and poly == Poly(
        2*x**2 + 3, x, modulus=5) and poly.get_domain() == FF(5)

    coeff, poly = Poly(x/2 + 1, x).clear_denoms(convert=True)
    assert coeff == 2 and poly == Poly(
        x + 2, x, domain='ZZ') and poly.get_domain() == ZZ

    coeff, poly = Poly(x/y + 1, x).clear_denoms(convert=True)
    assert coeff == y and poly == Poly(
        x + y, x, domain='ZZ[y]') and poly.get_domain() == ZZ[y]

    coeff, poly = Poly(x/3 + sqrt(2), x, domain='EX').clear_denoms()
    assert coeff == 3 and poly == Poly(
        x + 3*sqrt(2), x, domain='EX') and poly.get_domain() == EX

    coeff, poly = Poly(
        x/3 + sqrt(2), x, domain='EX').clear_denoms(convert=True)
    assert coeff == 3 and poly == Poly(
        x + 3*sqrt(2), x, domain='EX') and poly.get_domain() == EX


def test_Poly_rat_clear_denoms():
    f = Poly(x**2/y + 1, x)
    g = Poly(x**3 + y, x)

    assert f.rat_clear_denoms(g) == \
        (Poly(x**2 + y, x), Poly(y*x**3 + y**2, x))

    f = f.set_domain(EX)
    g = g.set_domain(EX)

    assert f.rat_clear_denoms(g) == (f, g)


def test_issue_20427():
    f = Poly(-117968192370600*18**(S(1)/3)/(217603955769048*(24201 +
        253*sqrt(9165))**(S(1)/3) + 2273005839412*sqrt(9165)*(24201 +
        253*sqrt(9165))**(S(1)/3)) - 15720318185*2**(S(2)/3)*3**(S(1)/3)*(24201
        + 253*sqrt(9165))**(S(2)/3)/(217603955769048*(24201 + 253*sqrt(9165))**
        (S(1)/3) + 2273005839412*sqrt(9165)*(24201 + 253*sqrt(9165))**(S(1)/3))
        + 15720318185*12**(S(1)/3)*(24201 + 253*sqrt(9165))**(S(2)/3)/(
        217603955769048*(24201 + 253*sqrt(9165))**(S(1)/3) + 2273005839412*
        sqrt(9165)*(24201 + 253*sqrt(9165))**(S(1)/3)) + 117968192370600*2**(
        S(1)/3)*3**(S(2)/3)/(217603955769048*(24201 + 253*sqrt(9165))**(S(1)/3)
        + 2273005839412*sqrt(9165)*(24201 + 253*sqrt(9165))**(S(1)/3)), x)
    assert f == Poly(0, x, domain='EX')


def test_Poly_integrate():
    assert Poly(x + 1).integrate() == Poly(x**2/2 + x)
    assert Poly(x + 1).integrate(x) == Poly(x**2/2 + x)
    assert Poly(x + 1).integrate((x, 1)) == Poly(x**2/2 + x)

    assert Poly(x*y + 1).integrate(x) == Poly(x**2*y/2 + x)
    assert Poly(x*y + 1).integrate(y) == Poly(x*y**2/2 + y)

    assert Poly(x*y + 1).integrate(x, x) == Poly(x**3*y/6 + x**2/2)
    assert Poly(x*y + 1).integrate(y, y) == Poly(x*y**3/6 + y**2/2)

    assert Poly(x*y + 1).integrate((x, 2)) == Poly(x**3*y/6 + x**2/2)
    assert Poly(x*y + 1).integrate((y, 2)) == Poly(x*y**3/6 + y**2/2)

    assert Poly(x*y + 1).integrate(x, y) == Poly(x**2*y**2/4 + x*y)
    assert Poly(x*y + 1).integrate(y, x) == Poly(x**2*y**2/4 + x*y)


def test_Poly_diff():
    assert Poly(x**2 + x).diff() == Poly(2*x + 1)
    assert Poly(x**2 + x).diff(x) == Poly(2*x + 1)
    assert Poly(x**2 + x).diff((x, 1)) == Poly(2*x + 1)

    assert Poly(x**2*y**2 + x*y).diff(x) == Poly(2*x*y**2 + y)
    assert Poly(x**2*y**2 + x*y).diff(y) == Poly(2*x**2*y + x)

    assert Poly(x**2*y**2 + x*y).diff(x, x) == Poly(2*y**2, x, y)
    assert Poly(x**2*y**2 + x*y).diff(y, y) == Poly(2*x**2, x, y)

    assert Poly(x**2*y**2 + x*y).diff((x, 2)) == Poly(2*y**2, x, y)
    assert Poly(x**2*y**2 + x*y).diff((y, 2)) == Poly(2*x**2, x, y)

    assert Poly(x**2*y**2 + x*y).diff(x, y) == Poly(4*x*y + 1)
    assert Poly(x**2*y**2 + x*y).diff(y, x) == Poly(4*x*y + 1)


def test_issue_9585():
    assert diff(Poly(x**2 + x)) == Poly(2*x + 1)
    assert diff(Poly(x**2 + x), x, evaluate=False) == \
        Derivative(Poly(x**2 + x), x)
    assert Derivative(Poly(x**2 + x), x).doit() == Poly(2*x + 1)


def test_Poly_eval():
    assert Poly(0, x).eval(7) == 0
    assert Poly(1, x).eval(7) == 1
    assert Poly(x, x).eval(7) == 7

    assert Poly(0, x).eval(0, 7) == 0
    assert Poly(1, x).eval(0, 7) == 1
    assert Poly(x, x).eval(0, 7) == 7

    assert Poly(0, x).eval(x, 7) == 0
    assert Poly(1, x).eval(x, 7) == 1
    assert Poly(x, x).eval(x, 7) == 7

    assert Poly(0, x).eval('x', 7) == 0
    assert Poly(1, x).eval('x', 7) == 1
    assert Poly(x, x).eval('x', 7) == 7

    raises(PolynomialError, lambda: Poly(1, x).eval(1, 7))
    raises(PolynomialError, lambda: Poly(1, x).eval(y, 7))
    raises(PolynomialError, lambda: Poly(1, x).eval('y', 7))

    assert Poly(123, x, y).eval(7) == Poly(123, y)
    assert Poly(2*y, x, y).eval(7) == Poly(2*y, y)
    assert Poly(x*y, x, y).eval(7) == Poly(7*y, y)

    assert Poly(123, x, y).eval(x, 7) == Poly(123, y)
    assert Poly(2*y, x, y).eval(x, 7) == Poly(2*y, y)
    assert Poly(x*y, x, y).eval(x, 7) == Poly(7*y, y)

    assert Poly(123, x, y).eval(y, 7) == Poly(123, x)
    assert Poly(2*y, x, y).eval(y, 7) == Poly(14, x)
    assert Poly(x*y, x, y).eval(y, 7) == Poly(7*x, x)

    assert Poly(x*y + y, x, y).eval({x: 7}) == Poly(8*y, y)
    assert Poly(x*y + y, x, y).eval({y: 7}) == Poly(7*x + 7, x)

    assert Poly(x*y + y, x, y).eval({x: 6, y: 7}) == 49
    assert Poly(x*y + y, x, y).eval({x: 7, y: 6}) == 48

    assert Poly(x*y + y, x, y).eval((6, 7)) == 49
    assert Poly(x*y + y, x, y).eval([6, 7]) == 49

    assert Poly(x + 1, domain='ZZ').eval(S.Half) == Rational(3, 2)
    assert Poly(x + 1, domain='ZZ').eval(sqrt(2)) == sqrt(2) + 1

    raises(ValueError, lambda: Poly(x*y + y, x, y).eval((6, 7, 8)))
    raises(DomainError, lambda: Poly(x + 1, domain='ZZ').eval(S.Half, auto=False))

    # issue 6344
    alpha = Symbol('alpha')
    result = (2*alpha*z - 2*alpha + z**2 + 3)/(z**2 - 2*z + 1)

    f = Poly(x**2 + (alpha - 1)*x - alpha + 1, x, domain='ZZ[alpha]')
    assert f.eval((z + 1)/(z - 1)) == result

    g = Poly(x**2 + (alpha - 1)*x - alpha + 1, x, y, domain='ZZ[alpha]')
    assert g.eval((z + 1)/(z - 1)) == Poly(result, y, domain='ZZ(alpha,z)')

def test_Poly___call__():
    f = Poly(2*x*y + 3*x + y + 2*z)

    assert f(2) == Poly(5*y + 2*z + 6)
    assert f(2, 5) == Poly(2*z + 31)
    assert f(2, 5, 7) == 45


def test_parallel_poly_from_expr():
    assert parallel_poly_from_expr(
        [x - 1, x**2 - 1], x)[0] == [Poly(x - 1, x), Poly(x**2 - 1, x)]
    assert parallel_poly_from_expr(
        [Poly(x - 1, x), x**2 - 1], x)[0] == [Poly(x - 1, x), Poly(x**2 - 1, x)]
    assert parallel_poly_from_expr(
        [x - 1, Poly(x**2 - 1, x)], x)[0] == [Poly(x - 1, x), Poly(x**2 - 1, x)]
    assert parallel_poly_from_expr([Poly(
        x - 1, x), Poly(x**2 - 1, x)], x)[0] == [Poly(x - 1, x), Poly(x**2 - 1, x)]

    assert parallel_poly_from_expr(
        [x - 1, x**2 - 1], x, y)[0] == [Poly(x - 1, x, y), Poly(x**2 - 1, x, y)]
    assert parallel_poly_from_expr([Poly(
        x - 1, x), x**2 - 1], x, y)[0] == [Poly(x - 1, x, y), Poly(x**2 - 1, x, y)]
    assert parallel_poly_from_expr([x - 1, Poly(
        x**2 - 1, x)], x, y)[0] == [Poly(x - 1, x, y), Poly(x**2 - 1, x, y)]
    assert parallel_poly_from_expr([Poly(x - 1, x), Poly(
        x**2 - 1, x)], x, y)[0] == [Poly(x - 1, x, y), Poly(x**2 - 1, x, y)]

    assert parallel_poly_from_expr(
        [x - 1, x**2 - 1])[0] == [Poly(x - 1, x), Poly(x**2 - 1, x)]
    assert parallel_poly_from_expr(
        [Poly(x - 1, x), x**2 - 1])[0] == [Poly(x - 1, x), Poly(x**2 - 1, x)]
    assert parallel_poly_from_expr(
        [x - 1, Poly(x**2 - 1, x)])[0] == [Poly(x - 1, x), Poly(x**2 - 1, x)]
    assert parallel_poly_from_expr(
        [Poly(x - 1, x), Poly(x**2 - 1, x)])[0] == [Poly(x - 1, x), Poly(x**2 - 1, x)]

    assert parallel_poly_from_expr(
        [1, x**2 - 1])[0] == [Poly(1, x), Poly(x**2 - 1, x)]
    assert parallel_poly_from_expr(
        [1, x**2 - 1])[0] == [Poly(1, x), Poly(x**2 - 1, x)]
    assert parallel_poly_from_expr(
        [1, Poly(x**2 - 1, x)])[0] == [Poly(1, x), Poly(x**2 - 1, x)]
    assert parallel_poly_from_expr(
        [1, Poly(x**2 - 1, x)])[0] == [Poly(1, x), Poly(x**2 - 1, x)]

    assert parallel_poly_from_expr(
        [x**2 - 1, 1])[0] == [Poly(x**2 - 1, x), Poly(1, x)]
    assert parallel_poly_from_expr(
        [x**2 - 1, 1])[0] == [Poly(x**2 - 1, x), Poly(1, x)]
    assert parallel_poly_from_expr(
        [Poly(x**2 - 1, x), 1])[0] == [Poly(x**2 - 1, x), Poly(1, x)]
    assert parallel_poly_from_expr(
        [Poly(x**2 - 1, x), 1])[0] == [Poly(x**2 - 1, x), Poly(1, x)]

    assert parallel_poly_from_expr([Poly(x, x, y), Poly(y, x, y)], x, y, order='lex')[0] == \
        [Poly(x, x, y, domain='ZZ'), Poly(y, x, y, domain='ZZ')]

    raises(PolificationFailed, lambda: parallel_poly_from_expr([0, 1]))


def test_pdiv():
    f, g = x**2 - y**2, x - y
    q, r = x + y, 0

    F, G, Q, R = [ Poly(h, x, y) for h in (f, g, q, r) ]

    assert F.pdiv(G) == (Q, R)
    assert F.prem(G) == R
    assert F.pquo(G) == Q
    assert F.pexquo(G) == Q

    assert pdiv(f, g) == (q, r)
    assert prem(f, g) == r
    assert pquo(f, g) == q
    assert pexquo(f, g) == q

    assert pdiv(f, g, x, y) == (q, r)
    assert prem(f, g, x, y) == r
    assert pquo(f, g, x, y) == q
    assert pexquo(f, g, x, y) == q

    assert pdiv(f, g, (x, y)) == (q, r)
    assert prem(f, g, (x, y)) == r
    assert pquo(f, g, (x, y)) == q
    assert pexquo(f, g, (x, y)) == q

    assert pdiv(F, G) == (Q, R)
    assert prem(F, G) == R
    assert pquo(F, G) == Q
    assert pexquo(F, G) == Q

    assert pdiv(f, g, polys=True) == (Q, R)
    assert prem(f, g, polys=True) == R
    assert pquo(f, g, polys=True) == Q
    assert pexquo(f, g, polys=True) == Q

    assert pdiv(F, G, polys=False) == (q, r)
    assert prem(F, G, polys=False) == r
    assert pquo(F, G, polys=False) == q
    assert pexquo(F, G, polys=False) == q

    raises(ComputationFailed, lambda: pdiv(4, 2))
    raises(ComputationFailed, lambda: prem(4, 2))
    raises(ComputationFailed, lambda: pquo(4, 2))
    raises(ComputationFailed, lambda: pexquo(4, 2))


def test_div():
    f, g = x**2 - y**2, x - y
    q, r = x + y, 0

    F, G, Q, R = [ Poly(h, x, y) for h in (f, g, q, r) ]

    assert F.div(G) == (Q, R)
    assert F.rem(G) == R
    assert F.quo(G) == Q
    assert F.exquo(G) == Q

    assert div(f, g) == (q, r)
    assert rem(f, g) == r
    assert quo(f, g) == q
    assert exquo(f, g) == q

    assert div(f, g, x, y) == (q, r)
    assert rem(f, g, x, y) == r
    assert quo(f, g, x, y) == q
    assert exquo(f, g, x, y) == q

    assert div(f, g, (x, y)) == (q, r)
    assert rem(f, g, (x, y)) == r
    assert quo(f, g, (x, y)) == q
    assert exquo(f, g, (x, y)) == q

    assert div(F, G) == (Q, R)
    assert rem(F, G) == R
    assert quo(F, G) == Q
    assert exquo(F, G) == Q

    assert div(f, g, polys=True) == (Q, R)
    assert rem(f, g, polys=True) == R
    assert quo(f, g, polys=True) == Q
    assert exquo(f, g, polys=True) == Q

    assert div(F, G, polys=False) == (q, r)
    assert rem(F, G, polys=False) == r
    assert quo(F, G, polys=False) == q
    assert exquo(F, G, polys=False) == q

    raises(ComputationFailed, lambda: div(4, 2))
    raises(ComputationFailed, lambda: rem(4, 2))
    raises(ComputationFailed, lambda: quo(4, 2))
    raises(ComputationFailed, lambda: exquo(4, 2))

    f, g = x**2 + 1, 2*x - 4

    qz, rz = 0, x**2 + 1
    qq, rq = x/2 + 1, 5

    assert div(f, g) == (qq, rq)
    assert div(f, g, auto=True) == (qq, rq)
    assert div(f, g, auto=False) == (qz, rz)
    assert div(f, g, domain=ZZ) == (qz, rz)
    assert div(f, g, domain=QQ) == (qq, rq)
    assert div(f, g, domain=ZZ, auto=True) == (qq, rq)
    assert div(f, g, domain=ZZ, auto=False) == (qz, rz)
    assert div(f, g, domain=QQ, auto=True) == (qq, rq)
    assert div(f, g, domain=QQ, auto=False) == (qq, rq)

    assert rem(f, g) == rq
    assert rem(f, g, auto=True) == rq
    assert rem(f, g, auto=False) == rz
    assert rem(f, g, domain=ZZ) == rz
    assert rem(f, g, domain=QQ) == rq
    assert rem(f, g, domain=ZZ, auto=True) == rq
    assert rem(f, g, domain=ZZ, auto=False) == rz
    assert rem(f, g, domain=QQ, auto=True) == rq
    assert rem(f, g, domain=QQ, auto=False) == rq

    assert quo(f, g) == qq
    assert quo(f, g, auto=True) == qq
    assert quo(f, g, auto=False) == qz
    assert quo(f, g, domain=ZZ) == qz
    assert quo(f, g, domain=QQ) == qq
    assert quo(f, g, domain=ZZ, auto=True) == qq
    assert quo(f, g, domain=ZZ, auto=False) == qz
    assert quo(f, g, domain=QQ, auto=True) == qq
    assert quo(f, g, domain=QQ, auto=False) == qq

    f, g, q = x**2, 2*x, x/2

    assert exquo(f, g) == q
    assert exquo(f, g, auto=True) == q
    raises(ExactQuotientFailed, lambda: exquo(f, g, auto=False))
    raises(ExactQuotientFailed, lambda: exquo(f, g, domain=ZZ))
    assert exquo(f, g, domain=QQ) == q
    assert exquo(f, g, domain=ZZ, auto=True) == q
    raises(ExactQuotientFailed, lambda: exquo(f, g, domain=ZZ, auto=False))
    assert exquo(f, g, domain=QQ, auto=True) == q
    assert exquo(f, g, domain=QQ, auto=False) == q

    f, g = Poly(x**2), Poly(x)

    q, r = f.div(g)
    assert q.get_domain().is_ZZ and r.get_domain().is_ZZ
    r = f.rem(g)
    assert r.get_domain().is_ZZ
    q = f.quo(g)
    assert q.get_domain().is_ZZ
    q = f.exquo(g)
    assert q.get_domain().is_ZZ

    f, g = Poly(x+y, x), Poly(2*x+y, x)
    q, r = f.div(g)
    assert q.get_domain().is_Frac and r.get_domain().is_Frac

    # https://github.com/sympy/sympy/issues/19579
    p = Poly(2+3*I, x, domain=ZZ_I)
    q = Poly(1-I, x, domain=ZZ_I)
    assert p.div(q, auto=False) == \
        (Poly(0, x, domain='ZZ_I'), Poly(2 + 3*I, x, domain='ZZ_I'))
    assert p.div(q, auto=True) == \
        (Poly(-S(1)/2 + 5*I/2, x, domain='QQ_I'), Poly(0, x, domain='QQ_I'))

    f = 5*x**2 + 10*x + 3
    g = 2*x + 2
    assert div(f, g, domain=ZZ) == (0, f)


def test_issue_7864():
    q, r = div(a, .408248290463863*a)
    assert abs(q - 2.44948974278318) < 1e-14
    assert r == 0


def test_gcdex():
    f, g = 2*x, x**2 - 16
    s, t, h = x/32, Rational(-1, 16), 1

    F, G, S, T, H = [ Poly(u, x, domain='QQ') for u in (f, g, s, t, h) ]

    assert F.half_gcdex(G) == (S, H)
    assert F.gcdex(G) == (S, T, H)
    assert F.invert(G) == S

    assert half_gcdex(f, g) == (s, h)
    assert gcdex(f, g) == (s, t, h)
    assert invert(f, g) == s

    assert half_gcdex(f, g, x) == (s, h)
    assert gcdex(f, g, x) == (s, t, h)
    assert invert(f, g, x) == s

    assert half_gcdex(f, g, (x,)) == (s, h)
    assert gcdex(f, g, (x,)) == (s, t, h)
    assert invert(f, g, (x,)) == s

    assert half_gcdex(F, G) == (S, H)
    assert gcdex(F, G) == (S, T, H)
    assert invert(F, G) == S

    assert half_gcdex(f, g, polys=True) == (S, H)
    assert gcdex(f, g, polys=True) == (S, T, H)
    assert invert(f, g, polys=True) == S

    assert half_gcdex(F, G, polys=False) == (s, h)
    assert gcdex(F, G, polys=False) == (s, t, h)
    assert invert(F, G, polys=False) == s

    assert half_gcdex(100, 2004) == (-20, 4)
    assert gcdex(100, 2004) == (-20, 1, 4)
    assert invert(3, 7) == 5

    raises(DomainError, lambda: half_gcdex(x + 1, 2*x + 1, auto=False))
    raises(DomainError, lambda: gcdex(x + 1, 2*x + 1, auto=False))
    raises(DomainError, lambda: invert(x + 1, 2*x + 1, auto=False))


def test_revert():
    f = Poly(1 - x**2/2 + x**4/24 - x**6/720)
    g = Poly(61*x**6/720 + 5*x**4/24 + x**2/2 + 1)

    assert f.revert(8) == g


def test_subresultants():
    f, g, h = x**2 - 2*x + 1, x**2 - 1, 2*x - 2
    F, G, H = Poly(f), Poly(g), Poly(h)

    assert F.subresultants(G) == [F, G, H]
    assert subresultants(f, g) == [f, g, h]
    assert subresultants(f, g, x) == [f, g, h]
    assert subresultants(f, g, (x,)) == [f, g, h]
    assert subresultants(F, G) == [F, G, H]
    assert subresultants(f, g, polys=True) == [F, G, H]
    assert subresultants(F, G, polys=False) == [f, g, h]

    raises(ComputationFailed, lambda: subresultants(4, 2))


def test_resultant():
    f, g, h = x**2 - 2*x + 1, x**2 - 1, 0
    F, G = Poly(f), Poly(g)

    assert F.resultant(G) == h
    assert resultant(f, g) == h
    assert resultant(f, g, x) == h
    assert resultant(f, g, (x,)) == h
    assert resultant(F, G) == h
    assert resultant(f, g, polys=True) == h
    assert resultant(F, G, polys=False) == h
    assert resultant(f, g, includePRS=True) == (h, [f, g, 2*x - 2])

    f, g, h = x - a, x - b, a - b
    F, G, H = Poly(f), Poly(g), Poly(h)

    assert F.resultant(G) == H
    assert resultant(f, g) == h
    assert resultant(f, g, x) == h
    assert resultant(f, g, (x,)) == h
    assert resultant(F, G) == H
    assert resultant(f, g, polys=True) == H
    assert resultant(F, G, polys=False) == h

    raises(ComputationFailed, lambda: resultant(4, 2))


def test_discriminant():
    f, g = x**3 + 3*x**2 + 9*x - 13, -11664
    F = Poly(f)

    assert F.discriminant() == g
    assert discriminant(f) == g
    assert discriminant(f, x) == g
    assert discriminant(f, (x,)) == g
    assert discriminant(F) == g
    assert discriminant(f, polys=True) == g
    assert discriminant(F, polys=False) == g

    f, g = a*x**2 + b*x + c, b**2 - 4*a*c
    F, G = Poly(f), Poly(g)

    assert F.discriminant() == G
    assert discriminant(f) == g
    assert discriminant(f, x, a, b, c) == g
    assert discriminant(f, (x, a, b, c)) == g
    assert discriminant(F) == G
    assert discriminant(f, polys=True) == G
    assert discriminant(F, polys=False) == g

    raises(ComputationFailed, lambda: discriminant(4))


def test_dispersion():
    # We test only the API here. For more mathematical
    # tests see the dedicated test file.
    fp = poly((x + 1)*(x + 2), x)
    assert sorted(fp.dispersionset()) == [0, 1]
    assert fp.dispersion() == 1

    fp = poly(x**4 - 3*x**2 + 1, x)
    gp = fp.shift(-3)
    assert sorted(fp.dispersionset(gp)) == [2, 3, 4]
    assert fp.dispersion(gp) == 4


def test_gcd_list():
    F = [x**3 - 1, x**2 - 1, x**2 - 3*x + 2]

    assert gcd_list(F) == x - 1
    assert gcd_list(F, polys=True) == Poly(x - 1)

    assert gcd_list([]) == 0
    assert gcd_list([1, 2]) == 1
    assert gcd_list([4, 6, 8]) == 2

    assert gcd_list([x*(y + 42) - x*y - x*42]) == 0

    gcd = gcd_list([], x)
    assert gcd.is_Number and gcd is S.Zero

    gcd = gcd_list([], x, polys=True)
    assert gcd.is_Poly and gcd.is_zero

    a = sqrt(2)
    assert gcd_list([a, -a]) == gcd_list([-a, a]) == a

    raises(ComputationFailed, lambda: gcd_list([], polys=True))


def test_lcm_list():
    F = [x**3 - 1, x**2 - 1, x**2 - 3*x + 2]

    assert lcm_list(F) == x**5 - x**4 - 2*x**3 - x**2 + x + 2
    assert lcm_list(F, polys=True) == Poly(x**5 - x**4 - 2*x**3 - x**2 + x + 2)

    assert lcm_list([]) == 1
    assert lcm_list([1, 2]) == 2
    assert lcm_list([4, 6, 8]) == 24

    assert lcm_list([x*(y + 42) - x*y - x*42]) == 0

    lcm = lcm_list([], x)
    assert lcm.is_Number and lcm is S.One

    lcm = lcm_list([], x, polys=True)
    assert lcm.is_Poly and lcm.is_one

    raises(ComputationFailed, lambda: lcm_list([], polys=True))


def test_gcd():
    f, g = x**3 - 1, x**2 - 1
    s, t = x**2 + x + 1, x + 1
    h, r = x - 1, x**4 + x**3 - x - 1

    F, G, S, T, H, R = [ Poly(u) for u in (f, g, s, t, h, r) ]

    assert F.cofactors(G) == (H, S, T)
    assert F.gcd(G) == H
    assert F.lcm(G) == R

    assert cofactors(f, g) == (h, s, t)
    assert gcd(f, g) == h
    assert lcm(f, g) == r

    assert cofactors(f, g, x) == (h, s, t)
    assert gcd(f, g, x) == h
    assert lcm(f, g, x) == r

    assert cofactors(f, g, (x,)) == (h, s, t)
    assert gcd(f, g, (x,)) == h
    assert lcm(f, g, (x,)) == r

    assert cofactors(F, G) == (H, S, T)
    assert gcd(F, G) == H
    assert lcm(F, G) == R

    assert cofactors(f, g, polys=True) == (H, S, T)
    assert gcd(f, g, polys=True) == H
    assert lcm(f, g, polys=True) == R

    assert cofactors(F, G, polys=False) == (h, s, t)
    assert gcd(F, G, polys=False) == h
    assert lcm(F, G, polys=False) == r

    f, g = 1.0*x**2 - 1.0, 1.0*x - 1.0
    h, s, t = g, 1.0*x + 1.0, 1.0

    assert cofactors(f, g) == (h, s, t)
    assert gcd(f, g) == h
    assert lcm(f, g) == f

    f, g = 1.0*x**2 - 1.0, 1.0*x - 1.0
    h, s, t = g, 1.0*x + 1.0, 1.0

    assert cofactors(f, g) == (h, s, t)
    assert gcd(f, g) == h
    assert lcm(f, g) == f

    assert cofactors(8, 6) == (2, 4, 3)
    assert gcd(8, 6) == 2
    assert lcm(8, 6) == 24

    f, g = x**2 - 3*x - 4, x**3 - 4*x**2 + x - 4
    l = x**4 - 3*x**3 - 3*x**2 - 3*x - 4
    h, s, t = x - 4, x + 1, x**2 + 1

    assert cofactors(f, g, modulus=11) == (h, s, t)
    assert gcd(f, g, modulus=11) == h
    assert lcm(f, g, modulus=11) == l

    f, g = x**2 + 8*x + 7, x**3 + 7*x**2 + x + 7
    l = x**4 + 8*x**3 + 8*x**2 + 8*x + 7
    h, s, t = x + 7, x + 1, x**2 + 1

    assert cofactors(f, g, modulus=11, symmetric=False) == (h, s, t)
    assert gcd(f, g, modulus=11, symmetric=False) == h
    assert lcm(f, g, modulus=11, symmetric=False) == l

    a, b = sqrt(2), -sqrt(2)
    assert gcd(a, b) == gcd(b, a) == sqrt(2)

    a, b = sqrt(-2), -sqrt(-2)
    assert gcd(a, b) == gcd(b, a) == sqrt(2)

    assert gcd(Poly(x - 2, x), Poly(I*x, x)) == Poly(1, x, domain=ZZ_I)

    raises(TypeError, lambda: gcd(x))
    raises(TypeError, lambda: lcm(x))

    f = Poly(-1, x)
    g = Poly(1, x)
    assert lcm(f, g) == Poly(1, x)

    f = Poly(0, x)
    g = Poly([1, 1], x)
    for i in (f, g):
        assert lcm(i, 0) == 0
        assert lcm(0, i) == 0
        assert lcm(i, f) == 0
        assert lcm(f, i) == 0

    f = 4*x**2 + x + 2
    pfz = Poly(f, domain=ZZ)
    pfq = Poly(f, domain=QQ)

    assert pfz.gcd(pfz) == pfz
    assert pfz.lcm(pfz) == pfz
    assert pfq.gcd(pfq) == pfq.monic()
    assert pfq.lcm(pfq) == pfq.monic()
    assert gcd(f, f) == f
    assert lcm(f, f) == f
    assert gcd(f, f, domain=QQ) == monic(f)
    assert lcm(f, f, domain=QQ) == monic(f)


def test_gcd_numbers_vs_polys():
    assert isinstance(gcd(3, 9), Integer)
    assert isinstance(gcd(3*x, 9), Integer)

    assert gcd(3, 9) == 3
    assert gcd(3*x, 9) == 3

    assert isinstance(gcd(Rational(3, 2), Rational(9, 4)), Rational)
    assert isinstance(gcd(Rational(3, 2)*x, Rational(9, 4)), Rational)

    assert gcd(Rational(3, 2), Rational(9, 4)) == Rational(3, 4)
    assert gcd(Rational(3, 2)*x, Rational(9, 4)) == 1

    assert isinstance(gcd(3.0, 9.0), Float)
    assert isinstance(gcd(3.0*x, 9.0), Float)

    assert gcd(3.0, 9.0) == 1.0
    assert gcd(3.0*x, 9.0) == 1.0

    # partial fix of 20597
    assert gcd(Mul(2, 3, evaluate=False), 2) == 2


def test_terms_gcd():
    assert terms_gcd(1) == 1
    assert terms_gcd(1, x) == 1

    assert terms_gcd(x - 1) == x - 1
    assert terms_gcd(-x - 1) == -x - 1

    assert terms_gcd(2*x + 3) == 2*x + 3
    assert terms_gcd(6*x + 4) == Mul(2, 3*x + 2, evaluate=False)

    assert terms_gcd(x**3*y + x*y**3) == x*y*(x**2 + y**2)
    assert terms_gcd(2*x**3*y + 2*x*y**3) == 2*x*y*(x**2 + y**2)
    assert terms_gcd(x**3*y/2 + x*y**3/2) == x*y/2*(x**2 + y**2)

    assert terms_gcd(x**3*y + 2*x*y**3) == x*y*(x**2 + 2*y**2)
    assert terms_gcd(2*x**3*y + 4*x*y**3) == 2*x*y*(x**2 + 2*y**2)
    assert terms_gcd(2*x**3*y/3 + 4*x*y**3/5) == x*y*Rational(2, 15)*(5*x**2 + 6*y**2)

    assert terms_gcd(2.0*x**3*y + 4.1*x*y**3) == x*y*(2.0*x**2 + 4.1*y**2)
    assert _aresame(terms_gcd(2.0*x + 3), 2.0*x + 3)

    assert terms_gcd((3 + 3*x)*(x + x*y), expand=False) == \
        (3*x + 3)*(x*y + x)
    assert terms_gcd((3 + 3*x)*(x + x*sin(3 + 3*y)), expand=False, deep=True) == \
        3*x*(x + 1)*(sin(Mul(3, y + 1, evaluate=False)) + 1)
    assert terms_gcd(sin(x + x*y), deep=True) == \
        sin(x*(y + 1))

    eq = Eq(2*x, 2*y + 2*z*y)
    assert terms_gcd(eq) == Eq(2*x, 2*y*(z + 1))
    assert terms_gcd(eq, deep=True) == Eq(2*x, 2*y*(z + 1))

    raises(TypeError, lambda: terms_gcd(x < 2))


def test_trunc():
    f, g = x**5 + 2*x**4 + 3*x**3 + 4*x**2 + 5*x + 6, x**5 - x**4 + x**2 - x
    F, G = Poly(f), Poly(g)

    assert F.trunc(3) == G
    assert trunc(f, 3) == g
    assert trunc(f, 3, x) == g
    assert trunc(f, 3, (x,)) == g
    assert trunc(F, 3) == G
    assert trunc(f, 3, polys=True) == G
    assert trunc(F, 3, polys=False) == g

    f, g = 6*x**5 + 5*x**4 + 4*x**3 + 3*x**2 + 2*x + 1, -x**4 + x**3 - x + 1
    F, G = Poly(f), Poly(g)

    assert F.trunc(3) == G
    assert trunc(f, 3) == g
    assert trunc(f, 3, x) == g
    assert trunc(f, 3, (x,)) == g
    assert trunc(F, 3) == G
    assert trunc(f, 3, polys=True) == G
    assert trunc(F, 3, polys=False) == g

    f = Poly(x**2 + 2*x + 3, modulus=5)

    assert f.trunc(2) == Poly(x**2 + 1, modulus=5)


def test_monic():
    f, g = 2*x - 1, x - S.Half
    F, G = Poly(f, domain='QQ'), Poly(g)

    assert F.monic() == G
    assert monic(f) == g
    assert monic(f, x) == g
    assert monic(f, (x,)) == g
    assert monic(F) == G
    assert monic(f, polys=True) == G
    assert monic(F, polys=False) == g

    raises(ComputationFailed, lambda: monic(4))

    assert monic(2*x**2 + 6*x + 4, auto=False) == x**2 + 3*x + 2
    raises(ExactQuotientFailed, lambda: monic(2*x + 6*x + 1, auto=False))

    assert monic(2.0*x**2 + 6.0*x + 4.0) == 1.0*x**2 + 3.0*x + 2.0
    assert monic(2*x**2 + 3*x + 4, modulus=5) == x**2 - x + 2


def test_content():
    f, F = 4*x + 2, Poly(4*x + 2)

    assert F.content() == 2
    assert content(f) == 2

    raises(ComputationFailed, lambda: content(4))

    f = Poly(2*x, modulus=3)

    assert f.content() == 1


def test_primitive():
    f, g = 4*x + 2, 2*x + 1
    F, G = Poly(f), Poly(g)

    assert F.primitive() == (2, G)
    assert primitive(f) == (2, g)
    assert primitive(f, x) == (2, g)
    assert primitive(f, (x,)) == (2, g)
    assert primitive(F) == (2, G)
    assert primitive(f, polys=True) == (2, G)
    assert primitive(F, polys=False) == (2, g)

    raises(ComputationFailed, lambda: primitive(4))

    f = Poly(2*x, modulus=3)
    g = Poly(2.0*x, domain=RR)

    assert f.primitive() == (1, f)
    assert g.primitive() == (1.0, g)

    assert primitive(S('-3*x/4 + y + 11/8')) == \
        S('(1/8, -6*x + 8*y + 11)')


def test_compose():
    f = x**12 + 20*x**10 + 150*x**8 + 500*x**6 + 625*x**4 - 2*x**3 - 10*x + 9
    g = x**4 - 2*x + 9
    h = x**3 + 5*x

    F, G, H = map(Poly, (f, g, h))

    assert G.compose(H) == F
    assert compose(g, h) == f
    assert compose(g, h, x) == f
    assert compose(g, h, (x,)) == f
    assert compose(G, H) == F
    assert compose(g, h, polys=True) == F
    assert compose(G, H, polys=False) == f

    assert F.decompose() == [G, H]
    assert decompose(f) == [g, h]
    assert decompose(f, x) == [g, h]
    assert decompose(f, (x,)) == [g, h]
    assert decompose(F) == [G, H]
    assert decompose(f, polys=True) == [G, H]
    assert decompose(F, polys=False) == [g, h]

    raises(ComputationFailed, lambda: compose(4, 2))
    raises(ComputationFailed, lambda: decompose(4))

    assert compose(x**2 - y**2, x - y, x, y) == x**2 - 2*x*y
    assert compose(x**2 - y**2, x - y, y, x) == -y**2 + 2*x*y


def test_shift():
    assert Poly(x**2 - 2*x + 1, x).shift(2) == Poly(x**2 + 2*x + 1, x)


def test_shift_list():
    assert Poly(x*y, [x,y]).shift_list([1,2]) == Poly((x+1)*(y+2), [x,y])


def test_transform():
    # Also test that 3-way unification is done correctly
    assert Poly(x**2 - 2*x + 1, x).transform(Poly(x + 1), Poly(x - 1)) == \
        Poly(4, x) == \
        cancel((x - 1)**2*(x**2 - 2*x + 1).subs(x, (x + 1)/(x - 1)))

    assert Poly(x**2 - x/2 + 1, x).transform(Poly(x + 1), Poly(x - 1)) == \
        Poly(3*x**2/2 + Rational(5, 2), x) == \
        cancel((x - 1)**2*(x**2 - x/2 + 1).subs(x, (x + 1)/(x - 1)))

    assert Poly(x**2 - 2*x + 1, x).transform(Poly(x + S.Half), Poly(x - 1)) == \
        Poly(Rational(9, 4), x) == \
        cancel((x - 1)**2*(x**2 - 2*x + 1).subs(x, (x + S.Half)/(x - 1)))

    assert Poly(x**2 - 2*x + 1, x).transform(Poly(x + 1), Poly(x - S.Half)) == \
        Poly(Rational(9, 4), x) == \
        cancel((x - S.Half)**2*(x**2 - 2*x + 1).subs(x, (x + 1)/(x - S.Half)))

    # Unify ZZ, QQ, and RR
    assert Poly(x**2 - 2*x + 1, x).transform(Poly(x + 1.0), Poly(x - S.Half)) == \
        Poly(Rational(9, 4), x, domain='RR') == \
        cancel((x - S.Half)**2*(x**2 - 2*x + 1).subs(x, (x + 1.0)/(x - S.Half)))

    raises(ValueError, lambda: Poly(x*y).transform(Poly(x + 1), Poly(x - 1)))
    raises(ValueError, lambda: Poly(x).transform(Poly(y + 1), Poly(x - 1)))
    raises(ValueError, lambda: Poly(x).transform(Poly(x + 1), Poly(y - 1)))
    raises(ValueError, lambda: Poly(x).transform(Poly(x*y + 1), Poly(x - 1)))
    raises(ValueError, lambda: Poly(x).transform(Poly(x + 1), Poly(x*y - 1)))


def test_sturm():
    f, F = x, Poly(x, domain='QQ')
    g, G = 1, Poly(1, x, domain='QQ')

    assert F.sturm() == [F, G]
    assert sturm(f) == [f, g]
    assert sturm(f, x) == [f, g]
    assert sturm(f, (x,)) == [f, g]
    assert sturm(F) == [F, G]
    assert sturm(f, polys=True) == [F, G]
    assert sturm(F, polys=False) == [f, g]

    raises(ComputationFailed, lambda: sturm(4))
    raises(DomainError, lambda: sturm(f, auto=False))

    f = Poly(S(1024)/(15625*pi**8)*x**5
           - S(4096)/(625*pi**8)*x**4
           + S(32)/(15625*pi**4)*x**3
           - S(128)/(625*pi**4)*x**2
           + Rational(1, 62500)*x
           - Rational(1, 625), x, domain='ZZ(pi)')

    assert sturm(f) == \
        [Poly(x**3 - 100*x**2 + pi**4/64*x - 25*pi**4/16, x, domain='ZZ(pi)'),
         Poly(3*x**2 - 200*x + pi**4/64, x, domain='ZZ(pi)'),
         Poly((Rational(20000, 9) - pi**4/96)*x + 25*pi**4/18, x, domain='ZZ(pi)'),
         Poly((-3686400000000*pi**4 - 11520000*pi**8 - 9*pi**12)/(26214400000000 - 245760000*pi**4 + 576*pi**8), x, domain='ZZ(pi)')]


def test_gff():
    f = x**5 + 2*x**4 - x**3 - 2*x**2

    assert Poly(f).gff_list() == [(Poly(x), 1), (Poly(x + 2), 4)]
    assert gff_list(f) == [(x, 1), (x + 2, 4)]

    raises(NotImplementedError, lambda: gff(f))

    f = x*(x - 1)**3*(x - 2)**2*(x - 4)**2*(x - 5)

    assert Poly(f).gff_list() == [(
        Poly(x**2 - 5*x + 4), 1), (Poly(x**2 - 5*x + 4), 2), (Poly(x), 3)]
    assert gff_list(f) == [(x**2 - 5*x + 4, 1), (x**2 - 5*x + 4, 2), (x, 3)]

    raises(NotImplementedError, lambda: gff(f))


def test_norm():
    a, b = sqrt(2), sqrt(3)
    f = Poly(a*x + b*y, x, y, extension=(a, b))
    assert f.norm() == Poly(4*x**4 - 12*x**2*y**2 + 9*y**4, x, y, domain='QQ')


def test_sqf_norm():
    assert sqf_norm(x**2 - 2, extension=sqrt(3)) == \
        ([1], x**2 - 2*sqrt(3)*x + 1, x**4 - 10*x**2 + 1)
    assert sqf_norm(x**2 - 3, extension=sqrt(2)) == \
        ([1], x**2 - 2*sqrt(2)*x - 1, x**4 - 10*x**2 + 1)

    assert Poly(x**2 - 2, extension=sqrt(3)).sqf_norm() == \
        ([1], Poly(x**2 - 2*sqrt(3)*x + 1, x, extension=sqrt(3)),
              Poly(x**4 - 10*x**2 + 1, x, domain='QQ'))

    assert Poly(x**2 - 3, extension=sqrt(2)).sqf_norm() == \
        ([1], Poly(x**2 - 2*sqrt(2)*x - 1, x, extension=sqrt(2)),
              Poly(x**4 - 10*x**2 + 1, x, domain='QQ'))


def test_sqf():
    f = x**5 - x**3 - x**2 + 1
    g = x**3 + 2*x**2 + 2*x + 1
    h = x - 1

    p = x**4 + x**3 - x - 1

    F, G, H, P = map(Poly, (f, g, h, p))

    assert F.sqf_part() == P
    assert sqf_part(f) == p
    assert sqf_part(f, x) == p
    assert sqf_part(f, (x,)) == p
    assert sqf_part(F) == P
    assert sqf_part(f, polys=True) == P
    assert sqf_part(F, polys=False) == p

    assert F.sqf_list() == (1, [(G, 1), (H, 2)])
    assert sqf_list(f) == (1, [(g, 1), (h, 2)])
    assert sqf_list(f, x) == (1, [(g, 1), (h, 2)])
    assert sqf_list(f, (x,)) == (1, [(g, 1), (h, 2)])
    assert sqf_list(F) == (1, [(G, 1), (H, 2)])
    assert sqf_list(f, polys=True) == (1, [(G, 1), (H, 2)])
    assert sqf_list(F, polys=False) == (1, [(g, 1), (h, 2)])

    assert F.sqf_list_include() == [(G, 1), (H, 2)]

    raises(ComputationFailed, lambda: sqf_part(4))

    assert sqf(1) == 1
    assert sqf_list(1) == (1, [])

    assert sqf((2*x**2 + 2)**7) == 128*(x**2 + 1)**7

    assert sqf(f) == g*h**2
    assert sqf(f, x) == g*h**2
    assert sqf(f, (x,)) == g*h**2

    d = x**2 + y**2

    assert sqf(f/d) == (g*h**2)/d
    assert sqf(f/d, x) == (g*h**2)/d
    assert sqf(f/d, (x,)) == (g*h**2)/d

    assert sqf(x - 1) == x - 1
    assert sqf(-x - 1) == -x - 1

    assert sqf(x - 1) == x - 1
    assert sqf(6*x - 10) == Mul(2, 3*x - 5, evaluate=False)

    assert sqf((6*x - 10)/(3*x - 6)) == Rational(2, 3)*((3*x - 5)/(x - 2))
    assert sqf(Poly(x**2 - 2*x + 1)) == (x - 1)**2

    f = 3 + x - x*(1 + x) + x**2

    assert sqf(f) == 3

    f = (x**2 + 2*x + 1)**20000000000

    assert sqf(f) == (x + 1)**40000000000
    assert sqf_list(f) == (1, [(x + 1, 40000000000)])

    # https://github.com/sympy/sympy/issues/26497
    assert sqf(expand(((y - 2)**2 * (y + 2) * (x + 1)))) == \
        (y - 2)**2 * expand((y + 2) * (x + 1))
    assert sqf(expand(((y - 2)**2 * (y + 2) * (z + 1)))) == \
        (y - 2)**2 * expand((y + 2) * (z + 1))
    assert sqf(expand(((y - I)**2 * (y + I) * (x + 1)))) == \
        (y - I)**2 * expand((y + I) * (x + 1))
    assert sqf(expand(((y - I)**2 * (y + I) * (z + 1)))) == \
        (y - I)**2 * expand((y + I) * (z + 1))

    # Check that factors are combined and sorted.
    p = (x - 2)**2*(x - 1)*(x + y)**2*(y - 2)**2*(y - 1)
    assert Poly(p).sqf_list() == (1, [
        (Poly(x*y - x - y + 1), 1),
        (Poly(x**2*y - 2*x**2 + x*y**2 - 4*x*y + 4*x - 2*y**2 + 4*y), 2)
    ])


def test_factor():
    f = x**5 - x**3 - x**2 + 1

    u = x + 1
    v = x - 1
    w = x**2 + x + 1

    F, U, V, W = map(Poly, (f, u, v, w))

    assert F.factor_list() == (1, [(U, 1), (V, 2), (W, 1)])
    assert factor_list(f) == (1, [(u, 1), (v, 2), (w, 1)])
    assert factor_list(f, x) == (1, [(u, 1), (v, 2), (w, 1)])
    assert factor_list(f, (x,)) == (1, [(u, 1), (v, 2), (w, 1)])
    assert factor_list(F) == (1, [(U, 1), (V, 2), (W, 1)])
    assert factor_list(f, polys=True) == (1, [(U, 1), (V, 2), (W, 1)])
    assert factor_list(F, polys=False) == (1, [(u, 1), (v, 2), (w, 1)])

    assert F.factor_list_include() == [(U, 1), (V, 2), (W, 1)]

    assert factor_list(1) == (1, [])
    assert factor_list(6) == (6, [])
    assert factor_list(sqrt(3), x) == (sqrt(3), [])
    assert factor_list((-1)**x, x) == (1, [(-1, x)])
    assert factor_list((2*x)**y, x) == (1, [(2, y), (x, y)])
    assert factor_list(sqrt(x*y), x) == (1, [(x*y, S.Half)])

    assert factor(6) == 6 and factor(6).is_Integer

    assert factor_list(3*x) == (3, [(x, 1)])
    assert factor_list(3*x**2) == (3, [(x, 2)])

    assert factor(3*x) == 3*x
    assert factor(3*x**2) == 3*x**2

    assert factor((2*x**2 + 2)**7) == 128*(x**2 + 1)**7

    assert factor(f) == u*v**2*w
    assert factor(f, x) == u*v**2*w
    assert factor(f, (x,)) == u*v**2*w

    g, p, q, r = x**2 - y**2, x - y, x + y, x**2 + 1

    assert factor(f/g) == (u*v**2*w)/(p*q)
    assert factor(f/g, x) == (u*v**2*w)/(p*q)
    assert factor(f/g, (x,)) == (u*v**2*w)/(p*q)

    p = Symbol('p', positive=True)
    i = Symbol('i', integer=True)
    r = Symbol('r', real=True)

    assert factor(sqrt(x*y)).is_Pow is True

    assert factor(sqrt(3*x**2 - 3)) == sqrt(3)*sqrt((x - 1)*(x + 1))
    assert factor(sqrt(3*x**2 + 3)) == sqrt(3)*sqrt(x**2 + 1)

    assert factor((y*x**2 - y)**i) == y**i*(x - 1)**i*(x + 1)**i
    assert factor((y*x**2 + y)**i) == y**i*(x**2 + 1)**i

    assert factor((y*x**2 - y)**t) == (y*(x - 1)*(x + 1))**t
    assert factor((y*x**2 + y)**t) == (y*(x**2 + 1))**t

    f = sqrt(expand((r**2 + 1)*(p + 1)*(p - 1)*(p - 2)**3))
    g = sqrt((p - 2)**3*(p - 1))*sqrt(p + 1)*sqrt(r**2 + 1)

    assert factor(f) == g
    assert factor(g) == g

    g = (x - 1)**5*(r**2 + 1)
    f = sqrt(expand(g))

    assert factor(f) == sqrt(g)

    f = Poly(sin(1)*x + 1, x, domain=EX)

    assert f.factor_list() == (1, [(f, 1)])

    f = x**4 + 1

    assert factor(f) == f
    assert factor(f, extension=I) == (x**2 - I)*(x**2 + I)
    assert factor(f, gaussian=True) == (x**2 - I)*(x**2 + I)
    assert factor(
        f, extension=sqrt(2)) == (x**2 + sqrt(2)*x + 1)*(x**2 - sqrt(2)*x + 1)

    assert factor(x**2 + 4*I*x - 4) == (x + 2*I)**2

    f = x**2 + 2*I*x - 4

    assert factor(f) == f

    f = 8192*x**2 + x*(22656 + 175232*I) - 921416 + 242313*I
    f_zzi = I*(x*(64 - 64*I) + 773 + 596*I)**2
    f_qqi = 8192*(x + S(177)/128 + 1369*I/128)**2

    assert factor(f) == f_zzi
    assert factor(f, domain=ZZ_I) == f_zzi
    assert factor(f, domain=QQ_I) == f_qqi

    f = x**2 + 2*sqrt(2)*x + 2

    assert factor(f, extension=sqrt(2)) == (x + sqrt(2))**2
    assert factor(f**3, extension=sqrt(2)) == (x + sqrt(2))**6

    assert factor(x**2 - 2*y**2, extension=sqrt(2)) == \
        (x + sqrt(2)*y)*(x - sqrt(2)*y)
    assert factor(2*x**2 - 4*y**2, extension=sqrt(2)) == \
        2*((x + sqrt(2)*y)*(x - sqrt(2)*y))

    assert factor(x - 1) == x - 1
    assert factor(-x - 1) == -x - 1

    assert factor(x - 1) == x - 1

    assert factor(6*x - 10) == Mul(2, 3*x - 5, evaluate=False)

    assert factor(x**11 + x + 1, modulus=65537, symmetric=True) == \
        (x**2 + x + 1)*(x**9 - x**8 + x**6 - x**5 + x**3 - x** 2 + 1)
    assert factor(x**11 + x + 1, modulus=65537, symmetric=False) == \
        (x**2 + x + 1)*(x**9 + 65536*x**8 + x**6 + 65536*x**5 +
         x**3 + 65536*x** 2 + 1)

    f = x/pi + x*sin(x)/pi
    g = y/(pi**2 + 2*pi + 1) + y*sin(x)/(pi**2 + 2*pi + 1)

    assert factor(f) == x*(sin(x) + 1)/pi
    assert factor(g) == y*(sin(x) + 1)/(pi + 1)**2

    assert factor(Eq(
        x**2 + 2*x + 1, x**3 + 1)) == Eq((x + 1)**2, (x + 1)*(x**2 - x + 1))

    f = (x**2 - 1)/(x**2 + 4*x + 4)

    assert factor(f) == (x + 1)*(x - 1)/(x + 2)**2
    assert factor(f, x) == (x + 1)*(x - 1)/(x + 2)**2

    f = 3 + x - x*(1 + x) + x**2

    assert factor(f) == 3
    assert factor(f, x) == 3

    assert factor(1/(x**2 + 2*x + 1/x) - 1) == -((1 - x + 2*x**2 +
                  x**3)/(1 + 2*x**2 + x**3))

    assert factor(f, expand=False) == f
    raises(PolynomialError, lambda: factor(f, x, expand=False))

    raises(FlagError, lambda: factor(x**2 - 1, polys=True))

    assert factor([x, Eq(x**2 - y**2, Tuple(x**2 - z**2, 1/x + 1/y))]) == \
        [x, Eq((x - y)*(x + y), Tuple((x - z)*(x + z), (x + y)/x/y))]

    assert not isinstance(
        Poly(x**3 + x + 1).factor_list()[1][0][0], PurePoly) is True
    assert isinstance(
        PurePoly(x**3 + x + 1).factor_list()[1][0][0], PurePoly) is True

    assert factor(sqrt(-x)) == sqrt(-x)

    # issue 5917
    e = (-2*x*(-x + 1)*(x - 1)*(-x*(-x + 1)*(x - 1) - x*(x - 1)**2)*(x**2*(x -
    1) - x*(x - 1) - x) - (-2*x**2*(x - 1)**2 - x*(-x + 1)*(-x*(-x + 1) +
    x*(x - 1)))*(x**2*(x - 1)**4 - x*(-x*(-x + 1)*(x - 1) - x*(x - 1)**2)))
    assert factor(e) == 0

    # deep option
    assert factor(sin(x**2 + x) + x, deep=True) == sin(x*(x + 1)) + x
    assert factor(sin(x**2 + x)*x, deep=True) == sin(x*(x + 1))*x

    assert factor(sqrt(x**2)) == sqrt(x**2)

    # issue 13149
    assert factor(expand((0.5*x+1)*(0.5*y+1))) == Mul(1.0, 0.5*x + 1.0,
        0.5*y + 1.0, evaluate = False)
    assert factor(expand((0.5*x+0.5)**2)) == 0.25*(1.0*x + 1.0)**2

    eq = x**2*y**2 + 11*x**2*y + 30*x**2 + 7*x*y**2 + 77*x*y + 210*x + 12*y**2 + 132*y + 360
    assert factor(eq, x) == (x + 3)*(x + 4)*(y**2 + 11*y + 30)
    assert factor(eq, x, deep=True) == (x + 3)*(x + 4)*(y**2 + 11*y + 30)
    assert factor(eq, y, deep=True) == (y + 5)*(y + 6)*(x**2 + 7*x + 12)

    # fraction option
    f = 5*x + 3*exp(2 - 7*x)
    assert factor(f, deep=True) == factor(f, deep=True, fraction=True)
    assert factor(f, deep=True, fraction=False) == 5*x + 3*exp(2)*exp(-7*x)

    assert factor_list(x**3 - x*y**2, t, w, x) == (
        1, [(x, 1), (x - y, 1), (x + y, 1)])
    assert factor_list((x+1)*(x**6-1)) == (
        1, [(x - 1, 1), (x + 1, 2), (x**2 - x + 1, 1), (x**2 + x + 1, 1)])

    # https://github.com/sympy/sympy/issues/24952
    s2, s2p, s2n = sqrt(2), 1 + sqrt(2), 1 - sqrt(2)
    pip, pin = 1 + pi, 1 - pi
    assert factor_list(s2p*s2n) == (-1, [(-s2n, 1), (s2p, 1)])
    assert factor_list(pip*pin) == (-1, [(-pin, 1), (pip, 1)])
    # Not sure about this one. Maybe coeff should be 1 or -1?
    assert factor_list(s2*s2n) == (-s2, [(-s2n, 1)])
    assert factor_list(pi*pin) == (-1, [(-pin, 1), (pi, 1)])
    assert factor_list(s2p*s2n, x) == (s2p*s2n, [])
    assert factor_list(pip*pin, x) == (pip*pin, [])
    assert factor_list(s2*s2n, x) == (s2*s2n, [])
    assert factor_list(pi*pin, x) == (pi*pin, [])
    assert factor_list((x - sqrt(2)*pi)*(x + sqrt(2)*pi), x) == (
        1, [(x - sqrt(2)*pi, 1), (x + sqrt(2)*pi, 1)])

    # https://github.com/sympy/sympy/issues/26497
    p = ((y - I)**2 * (y + I) * (x + 1))
    assert factor(expand(p)) == p

    p = ((x - I)**2 * (x + I) * (y + 1))
    assert factor(expand(p)) == p

    p = (y + 1)**2*(y + sqrt(2))**2*(x**2 + x + 2 + 3*sqrt(2))**2
    assert factor(expand(p), extension=True) == p

    e = (
        -x**2*y**4/(y**2 + 1) + 2*I*x**2*y**3/(y**2 + 1) + 2*I*x**2*y/(y**2 + 1) +
        x**2/(y**2 + 1) - 2*x*y**4/(y**2 + 1) + 4*I*x*y**3/(y**2 + 1) +
        4*I*x*y/(y**2 + 1) + 2*x/(y**2 + 1) - y**4 - y**4/(y**2 + 1) + 2*I*y**3
        + 2*I*y**3/(y**2 + 1) + 2*I*y + 2*I*y/(y**2 + 1) + 1 + 1/(y**2 + 1)
    )
    assert factor(e) == -(y - I)**3*(y + I)*(x**2 + 2*x + y**2 + 2)/(y**2 + 1)

    # issue 27506
    e = (I*t*x*y - 3*I*t - I*x*y*z - 6*x*y + 3*I*z + 18)
    assert factor(e) == -I*(x*y - 3)*(-t + z - 6*I)

    e = (8*x**2*z**2 - 32*x**2*z*t + 24*x**2*t**2 - 4*I*x*y*z**2 + 16*I*x*y*z*t -
         12*I*x*y*t**2 + z**4 - 8*z**3*t + 22*z**2*t**2 - 24*z*t**3 + 9*t**4)
    assert factor(e) == (-3*t + z)*(-t + z)*(3*t**2 - 4*t*z + 8*x**2 - 4*I*x*y + z**2)


def test_factor_large():
    f = (x**2 + 4*x + 4)**10000000*(x**2 + 1)*(x**2 + 2*x + 1)**1234567
    g = ((x**2 + 2*x + 1)**3000*y**2 + (x**2 + 2*x + 1)**3000*2*y + (
        x**2 + 2*x + 1)**3000)

    assert factor(f) == (x + 2)**20000000*(x**2 + 1)*(x + 1)**2469134
    assert factor(g) == (x + 1)**6000*(y + 1)**2

    assert factor_list(
        f) == (1, [(x + 1, 2469134), (x + 2, 20000000), (x**2 + 1, 1)])
    assert factor_list(g) == (1, [(y + 1, 2), (x + 1, 6000)])

    f = (x**2 - y**2)**200000*(x**7 + 1)
    g = (x**2 + y**2)**200000*(x**7 + 1)

    assert factor(f) == \
        (x + 1)*(x - y)**200000*(x + y)**200000*(x**6 - x**5 +
         x**4 - x**3 + x**2 - x + 1)
    assert factor(g, gaussian=True) == \
        (x + 1)*(x - I*y)**200000*(x + I*y)**200000*(x**6 - x**5 +
         x**4 - x**3 + x**2 - x + 1)

    assert factor_list(f) == \
        (1, [(x + 1, 1), (x - y, 200000), (x + y, 200000), (x**6 -
         x**5 + x**4 - x**3 + x**2 - x + 1, 1)])
    assert factor_list(g, gaussian=True) == \
        (1, [(x + 1, 1), (x - I*y, 200000), (x + I*y, 200000), (
            x**6 - x**5 + x**4 - x**3 + x**2 - x + 1, 1)])


def test_factor_noeval():
    assert factor(6*x - 10) == Mul(2, 3*x - 5, evaluate=False)
    assert factor((6*x - 10)/(3*x - 6)) == Mul(Rational(2, 3), 3*x - 5, 1/(x - 2))


def test_intervals():
    assert intervals(0) == []
    assert intervals(1) == []

    assert intervals(x, sqf=True) == [(0, 0)]
    assert intervals(x) == [((0, 0), 1)]

    assert intervals(x**128) == [((0, 0), 128)]
    assert intervals([x**2, x**4]) == [((0, 0), {0: 2, 1: 4})]

    f = Poly((x*Rational(2, 5) - Rational(17, 3))*(4*x + Rational(1, 257)))

    assert f.intervals(sqf=True) == [(-1, 0), (14, 15)]
    assert f.intervals() == [((-1, 0), 1), ((14, 15), 1)]

    assert f.intervals(fast=True, sqf=True) == [(-1, 0), (14, 15)]
    assert f.intervals(fast=True) == [((-1, 0), 1), ((14, 15), 1)]

    assert f.intervals(eps=Rational(1, 10)) == f.intervals(eps=0.1) == \
        [((Rational(-1, 258), 0), 1), ((Rational(85, 6), Rational(85, 6)), 1)]
    assert f.intervals(eps=Rational(1, 100)) == f.intervals(eps=0.01) == \
        [((Rational(-1, 258), 0), 1), ((Rational(85, 6), Rational(85, 6)), 1)]
    assert f.intervals(eps=Rational(1, 1000)) == f.intervals(eps=0.001) == \
        [((Rational(-1, 1002), 0), 1), ((Rational(85, 6), Rational(85, 6)), 1)]
    assert f.intervals(eps=Rational(1, 10000)) == f.intervals(eps=0.0001) == \
        [((Rational(-1, 1028), Rational(-1, 1028)), 1), ((Rational(85, 6), Rational(85, 6)), 1)]

    f = (x*Rational(2, 5) - Rational(17, 3))*(4*x + Rational(1, 257))

    assert intervals(f, sqf=True) == [(-1, 0), (14, 15)]
    assert intervals(f) == [((-1, 0), 1), ((14, 15), 1)]

    assert intervals(f, eps=Rational(1, 10)) == intervals(f, eps=0.1) == \
        [((Rational(-1, 258), 0), 1), ((Rational(85, 6), Rational(85, 6)), 1)]
    assert intervals(f, eps=Rational(1, 100)) == intervals(f, eps=0.01) == \
        [((Rational(-1, 258), 0), 1), ((Rational(85, 6), Rational(85, 6)), 1)]
    assert intervals(f, eps=Rational(1, 1000)) == intervals(f, eps=0.001) == \
        [((Rational(-1, 1002), 0), 1), ((Rational(85, 6), Rational(85, 6)), 1)]
    assert intervals(f, eps=Rational(1, 10000)) == intervals(f, eps=0.0001) == \
        [((Rational(-1, 1028), Rational(-1, 1028)), 1), ((Rational(85, 6), Rational(85, 6)), 1)]

    f = Poly((x**2 - 2)*(x**2 - 3)**7*(x + 1)*(7*x + 3)**3)

    assert f.intervals() == \
        [((-2, Rational(-3, 2)), 7), ((Rational(-3, 2), -1), 1),
         ((-1, -1), 1), ((-1, 0), 3),
         ((1, Rational(3, 2)), 1), ((Rational(3, 2), 2), 7)]

    assert intervals([x**5 - 200, x**5 - 201]) == \
        [((Rational(75, 26), Rational(101, 35)), {0: 1}), ((Rational(309, 107), Rational(26, 9)), {1: 1})]

    assert intervals([x**5 - 200, x**5 - 201], fast=True) == \
        [((Rational(75, 26), Rational(101, 35)), {0: 1}), ((Rational(309, 107), Rational(26, 9)), {1: 1})]

    assert intervals([x**2 - 200, x**2 - 201]) == \
        [((Rational(-71, 5), Rational(-85, 6)), {1: 1}), ((Rational(-85, 6), -14), {0: 1}),
         ((14, Rational(85, 6)), {0: 1}), ((Rational(85, 6), Rational(71, 5)), {1: 1})]

    assert intervals([x + 1, x + 2, x - 1, x + 1, 1, x - 1, x - 1, (x - 2)**2]) == \
        [((-2, -2), {1: 1}), ((-1, -1), {0: 1, 3: 1}), ((1, 1), {2:
          1, 5: 1, 6: 1}), ((2, 2), {7: 2})]

    f, g, h = x**2 - 2, x**4 - 4*x**2 + 4, x - 1

    assert intervals(f, inf=Rational(7, 4), sqf=True) == []
    assert intervals(f, inf=Rational(7, 5), sqf=True) == [(Rational(7, 5), Rational(3, 2))]
    assert intervals(f, sup=Rational(7, 4), sqf=True) == [(-2, -1), (1, Rational(3, 2))]
    assert intervals(f, sup=Rational(7, 5), sqf=True) == [(-2, -1)]

    assert intervals(g, inf=Rational(7, 4)) == []
    assert intervals(g, inf=Rational(7, 5)) == [((Rational(7, 5), Rational(3, 2)), 2)]
    assert intervals(g, sup=Rational(7, 4)) == [((-2, -1), 2), ((1, Rational(3, 2)), 2)]
    assert intervals(g, sup=Rational(7, 5)) == [((-2, -1), 2)]

    assert intervals([g, h], inf=Rational(7, 4)) == []
    assert intervals([g, h], inf=Rational(7, 5)) == [((Rational(7, 5), Rational(3, 2)), {0: 2})]
    assert intervals([g, h], sup=S(
        7)/4) == [((-2, -1), {0: 2}), ((1, 1), {1: 1}), ((1, Rational(3, 2)), {0: 2})]
    assert intervals(
        [g, h], sup=Rational(7, 5)) == [((-2, -1), {0: 2}), ((1, 1), {1: 1})]

    assert intervals([x + 2, x**2 - 2]) == \
        [((-2, -2), {0: 1}), ((-2, -1), {1: 1}), ((1, 2), {1: 1})]
    assert intervals([x + 2, x**2 - 2], strict=True) == \
        [((-2, -2), {0: 1}), ((Rational(-3, 2), -1), {1: 1}), ((1, 2), {1: 1})]

    f = 7*z**4 - 19*z**3 + 20*z**2 + 17*z + 20

    assert intervals(f) == []

    real_part, complex_part = intervals(f, all=True, sqf=True)

    assert real_part == []
    assert all(re(a) < re(r) < re(b) and im(
        a) < im(r) < im(b) for (a, b), r in zip(complex_part, nroots(f)))

    assert complex_part == [(Rational(-40, 7) - I*40/7, 0),
                            (Rational(-40, 7), I*40/7),
                            (I*Rational(-40, 7), Rational(40, 7)),
                            (0, Rational(40, 7) + I*40/7)]

    real_part, complex_part = intervals(f, all=True, sqf=True, eps=Rational(1, 10))

    assert real_part == []
    assert all(re(a) < re(r) < re(b) and im(
        a) < im(r) < im(b) for (a, b), r in zip(complex_part, nroots(f)))

    raises(ValueError, lambda: intervals(x**2 - 2, eps=10**-100000))
    raises(ValueError, lambda: Poly(x**2 - 2).intervals(eps=10**-100000))
    raises(
        ValueError, lambda: intervals([x**2 - 2, x**2 - 3], eps=10**-100000))


def test_refine_root():
    f = Poly(x**2 - 2)

    assert f.refine_root(1, 2, steps=0) == (1, 2)
    assert f.refine_root(-2, -1, steps=0) == (-2, -1)

    assert f.refine_root(1, 2, steps=None) == (1, Rational(3, 2))
    assert f.refine_root(-2, -1, steps=None) == (Rational(-3, 2), -1)

    assert f.refine_root(1, 2, steps=1) == (1, Rational(3, 2))
    assert f.refine_root(-2, -1, steps=1) == (Rational(-3, 2), -1)

    assert f.refine_root(1, 2, steps=1, fast=True) == (1, Rational(3, 2))
    assert f.refine_root(-2, -1, steps=1, fast=True) == (Rational(-3, 2), -1)

    assert f.refine_root(1, 2, eps=Rational(1, 100)) == (Rational(24, 17), Rational(17, 12))
    assert f.refine_root(1, 2, eps=1e-2) == (Rational(24, 17), Rational(17, 12))

    raises(PolynomialError, lambda: (f**2).refine_root(1, 2, check_sqf=True))

    raises(RefinementFailed, lambda: (f**2).refine_root(1, 2))
    raises(RefinementFailed, lambda: (f**2).refine_root(2, 3))

    f = x**2 - 2

    assert refine_root(f, 1, 2, steps=1) == (1, Rational(3, 2))
    assert refine_root(f, -2, -1, steps=1) == (Rational(-3, 2), -1)

    assert refine_root(f, 1, 2, steps=1, fast=True) == (1, Rational(3, 2))
    assert refine_root(f, -2, -1, steps=1, fast=True) == (Rational(-3, 2), -1)

    assert refine_root(f, 1, 2, eps=Rational(1, 100)) == (Rational(24, 17), Rational(17, 12))
    assert refine_root(f, 1, 2, eps=1e-2) == (Rational(24, 17), Rational(17, 12))

    raises(PolynomialError, lambda: refine_root(1, 7, 8, eps=Rational(1, 100)))

    raises(ValueError, lambda: Poly(f).refine_root(1, 2, eps=10**-100000))
    raises(ValueError, lambda: refine_root(f, 1, 2, eps=10**-100000))


def test_count_roots():
    assert count_roots(x**2 - 2) == 2

    assert count_roots(x**2 - 2, inf=-oo) == 2
    assert count_roots(x**2 - 2, sup=+oo) == 2
    assert count_roots(x**2 - 2, inf=-oo, sup=+oo) == 2

    assert count_roots(x**2 - 2, inf=-2) == 2
    assert count_roots(x**2 - 2, inf=-1) == 1

    assert count_roots(x**2 - 2, sup=1) == 1
    assert count_roots(x**2 - 2, sup=2) == 2

    assert count_roots(x**2 - 2, inf=-1, sup=1) == 0
    assert count_roots(x**2 - 2, inf=-2, sup=2) == 2

    assert count_roots(x**2 - 2, inf=-1, sup=1) == 0
    assert count_roots(x**2 - 2, inf=-2, sup=2) == 2

    assert count_roots(x**2 + 2) == 0
    assert count_roots(x**2 + 2, inf=-2*I) == 2
    assert count_roots(x**2 + 2, sup=+2*I) == 2
    assert count_roots(x**2 + 2, inf=-2*I, sup=+2*I) == 2

    assert count_roots(x**2 + 2, inf=0) == 0
    assert count_roots(x**2 + 2, sup=0) == 0

    assert count_roots(x**2 + 2, inf=-I) == 1
    assert count_roots(x**2 + 2, sup=+I) == 1

    assert count_roots(x**2 + 2, inf=+I/2, sup=+I) == 0
    assert count_roots(x**2 + 2, inf=-I, sup=-I/2) == 0

    raises(PolynomialError, lambda: count_roots(1))


def test_count_roots_extension():

    p1 = Poly(sqrt(2)*x**2 - 2, x, extension=True)
    assert p1.count_roots() == 2
    assert p1.count_roots(inf=0) == 1
    assert p1.count_roots(sup=0) == 1

    p2 = Poly(x**2 + sqrt(2), x, extension=True)
    assert p2.count_roots() == 0

    p3 = Poly(x**2 + 2*sqrt(2)*x + 1, x, extension=True)
    assert p3.count_roots() == 2
    assert p3.count_roots(inf=-10, sup=10) == 2
    assert p3.count_roots(inf=-10, sup=0) == 2
    assert p3.count_roots(inf=-10, sup=-3) == 0
    assert p3.count_roots(inf=-3, sup=-2) == 1
    assert p3.count_roots(inf=-1, sup=0) == 1


def test_Poly_root():
    f = Poly(2*x**3 - 7*x**2 + 4*x + 4)

    assert f.root(0) == Rational(-1, 2)
    assert f.root(1) == 2
    assert f.root(2) == 2
    raises(IndexError, lambda: f.root(3))

    assert Poly(x**5 + x + 1).root(0) == rootof(x**3 - x**2 + 1, 0)


def test_real_roots():

    assert real_roots(x) == [0]
    assert real_roots(x, multiple=False) == [(0, 1)]

    assert real_roots(x**3) == [0, 0, 0]
    assert real_roots(x**3, multiple=False) == [(0, 3)]

    assert real_roots(x*(x**3 + x + 3)) == [rootof(x**3 + x + 3, 0), 0]
    assert real_roots(x*(x**3 + x + 3), multiple=False) == [(rootof(
        x**3 + x + 3, 0), 1), (0, 1)]

    assert real_roots(
        x**3*(x**3 + x + 3)) == [rootof(x**3 + x + 3, 0), 0, 0, 0]
    assert real_roots(x**3*(x**3 + x + 3), multiple=False) == [(rootof(
        x**3 + x + 3, 0), 1), (0, 3)]

    assert real_roots(x**2 - 2, radicals=False) == [
            rootof(x**2 - 2, 0, radicals=False),
            rootof(x**2 - 2, 1, radicals=False),
        ]

    f = 2*x**3 - 7*x**2 + 4*x + 4
    g = x**3 + x + 1

    assert Poly(f).real_roots() == [Rational(-1, 2), 2, 2]
    assert Poly(g).real_roots() == [rootof(g, 0)]

    # testing extension
    f = x**2 - sqrt(2)
    roots = [-2**(S(1)/4), 2**(S(1)/4)]
    raises(NotImplementedError, lambda: real_roots(f))
    raises(NotImplementedError, lambda: real_roots(Poly(f, x)))
    assert real_roots(f, extension=True) == roots
    assert real_roots(Poly(f, extension=True)) == roots
    assert real_roots(Poly(f), extension=True) == roots


def test_all_roots():

    f = 2*x**3 - 7*x**2 + 4*x + 4
    froots = [Rational(-1, 2), 2, 2]
    assert all_roots(f) == Poly(f).all_roots() == froots

    g = x**3 + x + 1
    groots = [rootof(g, 0), rootof(g, 1), rootof(g, 2)]
    assert all_roots(g) == Poly(g).all_roots() == groots

    assert all_roots(x**2 - 2) == [-sqrt(2), sqrt(2)]
    assert all_roots(x**2 - 2, multiple=False) == [(-sqrt(2), 1), (sqrt(2), 1)]
    assert all_roots(x**2 - 2, radicals=False) == [
        rootof(x**2 - 2, 0, radicals=False),
        rootof(x**2 - 2, 1, radicals=False),
    ]

    p = x**5 - x - 1
    assert all_roots(p) == [
        rootof(p, 0), rootof(p, 1), rootof(p, 2), rootof(p, 3), rootof(p, 4)
    ]

    # testing extension
    f = x**2 + sqrt(2)
    roots = [-2**(S(1)/4)*I, 2**(S(1)/4)*I]
    raises(NotImplementedError, lambda: all_roots(f))
    raises(NotImplementedError, lambda : all_roots(Poly(f, x)))
    assert all_roots(f, extension=True) == roots
    assert all_roots(Poly(f, extension=True)) == roots
    assert all_roots(Poly(f), extension=True) == roots


def test_nroots():
    assert Poly(0, x).nroots() == []
    assert Poly(1, x).nroots() == []

    assert Poly(x**2 - 1, x).nroots() == [-1.0, 1.0]
    assert Poly(x**2 + 1, x).nroots() == [-1.0*I, 1.0*I]

    roots = Poly(x**2 - 1, x).nroots()
    assert roots == [-1.0, 1.0]

    roots = Poly(x**2 + 1, x).nroots()
    assert roots == [-1.0*I, 1.0*I]

    roots = Poly(x**2/3 - Rational(1, 3), x).nroots()
    assert roots == [-1.0, 1.0]

    roots = Poly(x**2/3 + Rational(1, 3), x).nroots()
    assert roots == [-1.0*I, 1.0*I]

    assert Poly(x**2 + 2*I, x).nroots() == [-1.0 + 1.0*I, 1.0 - 1.0*I]
    assert Poly(
        x**2 + 2*I, x, extension=I).nroots() == [-1.0 + 1.0*I, 1.0 - 1.0*I]

    assert Poly(0.2*x + 0.1).nroots() == [-0.5]

    roots = nroots(x**5 + x + 1, n=5)
    eps = Float("1e-5")

    assert re(roots[0]).epsilon_eq(-0.75487, eps) is S.true
    assert im(roots[0]) == 0
    assert re(roots[1]) == Float(-0.5, 5)
    assert im(roots[1]).epsilon_eq(-0.86602, eps) is S.true
    assert re(roots[2]) == Float(-0.5, 5)
    assert im(roots[2]).epsilon_eq(+0.86602, eps) is S.true
    assert re(roots[3]).epsilon_eq(+0.87743, eps) is S.true
    assert im(roots[3]).epsilon_eq(-0.74486, eps) is S.true
    assert re(roots[4]).epsilon_eq(+0.87743, eps) is S.true
    assert im(roots[4]).epsilon_eq(+0.74486, eps) is S.true

    eps = Float("1e-6")

    assert re(roots[0]).epsilon_eq(-0.75487, eps) is S.false
    assert im(roots[0]) == 0
    assert re(roots[1]) == Float(-0.5, 5)
    assert im(roots[1]).epsilon_eq(-0.86602, eps) is S.false
    assert re(roots[2]) == Float(-0.5, 5)
    assert im(roots[2]).epsilon_eq(+0.86602, eps) is S.false
    assert re(roots[3]).epsilon_eq(+0.87743, eps) is S.false
    assert im(roots[3]).epsilon_eq(-0.74486, eps) is S.false
    assert re(roots[4]).epsilon_eq(+0.87743, eps) is S.false
    assert im(roots[4]).epsilon_eq(+0.74486, eps) is S.false

    raises(DomainError, lambda: Poly(x + y, x).nroots())
    raises(MultivariatePolynomialError, lambda: Poly(x + y).nroots())

    assert nroots(x**2 - 1) == [-1.0, 1.0]

    roots = nroots(x**2 - 1)
    assert roots == [-1.0, 1.0]

    assert nroots(x + I) == [-1.0*I]
    assert nroots(x + 2*I) == [-2.0*I]

    raises(PolynomialError, lambda: nroots(0))

    # issue 8296
    f = Poly(x**4 - 1)
    assert f.nroots(2) == [w.n(2) for w in f.all_roots()]

    assert str(Poly(x**16 + 32*x**14 + 508*x**12 + 5440*x**10 +
        39510*x**8 + 204320*x**6 + 755548*x**4 + 1434496*x**2 +
        877969).nroots(2)) == ('[-1.7 - 1.9*I, -1.7 + 1.9*I, -1.7 '
        '- 2.5*I, -1.7 + 2.5*I, -1.0*I, 1.0*I, -1.7*I, 1.7*I, -2.8*I, '
        '2.8*I, -3.4*I, 3.4*I, 1.7 - 1.9*I, 1.7 + 1.9*I, 1.7 - 2.5*I, '
        '1.7 + 2.5*I]')
    assert str(Poly(1e-15*x**2 -1).nroots()) == ('[-31622776.6016838, 31622776.6016838]')

    # https://github.com/sympy/sympy/issues/23861

    i = Float('3.000000000000000000000000000000000000000000000000001')
    [r] = nroots(x + I*i, n=300)
    assert abs(r + I*i) < 1e-300


def test_ground_roots():
    f = x**6 - 4*x**4 + 4*x**3 - x**2

    assert Poly(f).ground_roots() == {S.One: 2, S.Zero: 2}
    assert ground_roots(f) == {S.One: 2, S.Zero: 2}


def test_nth_power_roots_poly():
    f = x**4 - x**2 + 1

    f_2 = (x**2 - x + 1)**2
    f_3 = (x**2 + 1)**2
    f_4 = (x**2 + x + 1)**2
    f_12 = (x - 1)**4

    assert nth_power_roots_poly(f, 1) == f

    raises(ValueError, lambda: nth_power_roots_poly(f, 0))
    raises(ValueError, lambda: nth_power_roots_poly(f, x))

    assert factor(nth_power_roots_poly(f, 2)) == f_2
    assert factor(nth_power_roots_poly(f, 3)) == f_3
    assert factor(nth_power_roots_poly(f, 4)) == f_4
    assert factor(nth_power_roots_poly(f, 12)) == f_12

    raises(MultivariatePolynomialError, lambda: nth_power_roots_poly(
        x + y, 2, x, y))

def test_which_real_roots():
    f = Poly(x**4 - 1)

    assert f.which_real_roots([1, -1]) == [1, -1]
    assert f.which_real_roots([1, -1, 2, 4]) == [1, -1]
    assert f.which_real_roots([1, -1, -1, 1, 2, 5]) == [1, -1]
    assert f.which_real_roots([10, 8, 7, -1, 1]) == [-1, 1]

    # no real roots
    # (technically its still a superset)
    f = Poly(x**2 + 1)
    assert f.which_real_roots([5, 10]) == []

    # not square free
    f = Poly((x-1)**2)
    assert f.which_real_roots([1, 1, -1, 2]) == [1]

    # candidates not superset
    f = Poly(x**2 - 1)
    assert f.which_real_roots([0, 2]) == [0, 2]

def test_which_all_roots():
    f = Poly(x**4 - 1)

    assert f.which_all_roots([1, -1, I, -I]) == [1, -1, I, -I]
    assert f.which_all_roots([I, I, -I, 1, -1]) == [I, -I, 1, -1]

    f = Poly(x**2 + 1)
    assert f.which_all_roots([I, -I, I/2]) == [I, -I]

    # not square free
    f = Poly((x-I)**2)
    assert f.which_all_roots([I, I, 1, -1, 0]) == [I]

    # candidates not superset
    f = Poly(x**2 + 1)
    assert f.which_all_roots([I/2, -I/2]) == [I/2, -I/2]

def test_same_root():
    f = Poly(x**4 + x**3 + x**2 + x + 1)
    eq = f.same_root
    r0 = exp(2 * I * pi / 5)
    assert [i for i, r in enumerate(f.all_roots()) if eq(r, r0)] == [3]

    raises(PolynomialError,
           lambda: Poly(x + 1, domain=QQ).same_root(0, 0))
    raises(DomainError,
           lambda: Poly(x**2 + 1, domain=FF(7)).same_root(0, 0))
    raises(DomainError,
           lambda: Poly(x ** 2 + 1, domain=ZZ_I).same_root(0, 0))
    raises(DomainError,
           lambda: Poly(y * x**2 + 1, domain=ZZ[y]).same_root(0, 0))
    raises(MultivariatePolynomialError,
           lambda: Poly(x * y + 1, domain=ZZ).same_root(0, 0))


def test_torational_factor_list():
    p = expand(((x**2-1)*(x-2)).subs({x:x*(1 + sqrt(2))}))
    assert _torational_factor_list(p, x) == (-2, [
        (-x*(1 + sqrt(2))/2 + 1, 1),
        (-x*(1 + sqrt(2)) - 1, 1),
        (-x*(1 + sqrt(2)) + 1, 1)])


    p = expand(((x**2-1)*(x-2)).subs({x:x*(1 + 2**Rational(1, 4))}))
    assert _torational_factor_list(p, x) is None


def test_cancel():
    assert cancel(0) == 0
    assert cancel(7) == 7
    assert cancel(x) == x

    assert cancel(oo) is oo

    raises(ValueError, lambda: cancel((1, 2, 3)))

    # test first tuple returnr
    assert (t:=cancel((2, 3))) == (1, 2, 3)
    assert isinstance(t, tuple)

    # tests 2nd tuple return
    assert (t:=cancel((1, 0), x)) == (1, 1, 0)
    assert isinstance(t, tuple)
    assert cancel((0, 1), x) == (1, 0, 1)

    f, g, p, q = 4*x**2 - 4, 2*x - 2, 2*x + 2, 1
    F, G, P, Q = [ Poly(u, x) for u in (f, g, p, q) ]

    assert F.cancel(G) == (1, P, Q)
    assert cancel((f, g)) == (1, p, q)
    assert cancel((f, g), x) == (1, p, q)
    assert cancel((f, g), (x,)) == (1, p, q)
    # tests 3rd tuple return
    assert (t:=cancel((F, G))) == (1, P, Q)
    assert isinstance(t, tuple)
    assert cancel((f, g), polys=True) == (1, P, Q)
    assert cancel((F, G), polys=False) == (1, p, q)

    f = (x**2 - 2)/(x + sqrt(2))

    assert cancel(f) == f
    assert cancel(f, greedy=False) == x - sqrt(2)

    f = (x**2 - 2)/(x - sqrt(2))

    assert cancel(f) == f
    assert cancel(f, greedy=False) == x + sqrt(2)

    assert cancel((x**2/4 - 1, x/2 - 1)) == (1, x + 2, 2)
    # assert cancel((x**2/4 - 1, x/2 - 1)) == (S.Half, x + 2, 1)

    assert cancel((x**2 - y)/(x - y)) == 1/(x - y)*(x**2 - y)

    assert cancel((x**2 - y**2)/(x - y), x) == x + y
    assert cancel((x**2 - y**2)/(x - y), y) == x + y
    assert cancel((x**2 - y**2)/(x - y)) == x + y

    assert cancel((x**3 - 1)/(x**2 - 1)) == (x**2 + x + 1)/(x + 1)
    assert cancel((x**3/2 - S.Half)/(x**2 - 1)) == (x**2 + x + 1)/(2*x + 2)

    assert cancel((exp(2*x) + 2*exp(x) + 1)/(exp(x) + 1)) == exp(x) + 1

    f = Poly(x**2 - a**2, x)
    g = Poly(x - a, x)

    F = Poly(x + a, x, domain='ZZ[a]')
    G = Poly(1, x, domain='ZZ[a]')

    assert cancel((f, g)) == (1, F, G)

    f = x**3 + (sqrt(2) - 2)*x**2 - (2*sqrt(2) + 3)*x - 3*sqrt(2)
    g = x**2 - 2

    assert cancel((f, g), extension=True) == (1, x**2 - 2*x - 3, x - sqrt(2))

    f = Poly(-2*x + 3, x)
    g = Poly(-x**9 + x**8 + x**6 - x**5 + 2*x**2 - 3*x + 1, x)

    assert cancel((f, g)) == (1, -f, -g)

    f = Poly(x/3 + 1, x)
    g = Poly(x/7 + 1, x)

    assert f.cancel(g) == (S(7)/3,
        Poly(x + 3, x, domain=QQ),
        Poly(x + 7, x, domain=QQ))
    assert f.cancel(g, include=True) == (
        Poly(7*x + 21, x, domain=QQ),
        Poly(3*x + 21, x, domain=QQ))

    pairs = [
        (1 + x, 1 + x, 1, 1, 1),
        (1 + x, 1 - x, -1, -1-x, -1+x),
        (1 - x, 1 + x, -1, 1-x, 1+x),
        (1 - x, 1 - x, 1, 1, 1),
    ]
    for f, g, coeff, p, q in pairs:
        assert cancel((f, g)) == (1, p, q)
        pf = Poly(f, x)
        pg = Poly(g, x)
        pp = Poly(p, x)
        pq = Poly(q, x)
        assert pf.cancel(pg) == (coeff, coeff*pp, pq)
        assert pf.rep.cancel(pg.rep) == (pp.rep, pq.rep)
        assert pf.rep.cancel(pg.rep, include=True) == (pp.rep, pq.rep)

    f = Poly(y, y, domain='ZZ(x)')
    g = Poly(1, y, domain='ZZ[x]')

    assert f.cancel(
        g) == (1, Poly(y, y, domain='ZZ(x)'), Poly(1, y, domain='ZZ(x)'))
    assert f.cancel(g, include=True) == (
        Poly(y, y, domain='ZZ(x)'), Poly(1, y, domain='ZZ(x)'))

    f = Poly(5*x*y + x, y, domain='ZZ(x)')
    g = Poly(2*x**2*y, y, domain='ZZ(x)')

    assert f.cancel(g, include=True) == (
        Poly(5*y + 1, y, domain='ZZ(x)'), Poly(2*x*y, y, domain='ZZ(x)'))

    f = -(-2*x - 4*y + 0.005*(z - y)**2)/((z - y)*(-z + y + 2))
    assert cancel(f).is_Mul == True

    P = tanh(x - 3.0)
    Q = tanh(x + 3.0)
    f = ((-2*P**2 + 2)*(-P**2 + 1)*Q**2/2 + (-2*P**2 + 2)*(-2*Q**2 + 2)*P*Q - (-2*P**2 + 2)*P**2*Q**2 + (-2*Q**2 + 2)*(-Q**2 + 1)*P**2/2 - (-2*Q**2 + 2)*P**2*Q**2)/(2*sqrt(P**2*Q**2 + 0.0001)) \
      + (-(-2*P**2 + 2)*P*Q**2/2 - (-2*Q**2 + 2)*P**2*Q/2)*((-2*P**2 + 2)*P*Q**2/2 + (-2*Q**2 + 2)*P**2*Q/2)/(2*(P**2*Q**2 + 0.0001)**Rational(3, 2))
    assert cancel(f).is_Mul == True

    # issue 7022
    A = Symbol('A', commutative=False)
    p1 = Piecewise((A*(x**2 - 1)/(x + 1), x > 1), ((x + 2)/(x**2 + 2*x), True))
    p2 = Piecewise((A*(x - 1), x > 1), (1/x, True))
    assert cancel(p1) == p2
    assert cancel(2*p1) == 2*p2
    assert cancel(1 + p1) == 1 + p2
    assert cancel((x**2 - 1)/(x + 1)*p1) == (x - 1)*p2
    assert cancel((x**2 - 1)/(x + 1) + p1) == (x - 1) + p2
    p3 = Piecewise(((x**2 - 1)/(x + 1), x > 1), ((x + 2)/(x**2 + 2*x), True))
    p4 = Piecewise(((x - 1), x > 1), (1/x, True))
    assert cancel(p3) == p4
    assert cancel(2*p3) == 2*p4
    assert cancel(1 + p3) == 1 + p4
    assert cancel((x**2 - 1)/(x + 1)*p3) == (x - 1)*p4
    assert cancel((x**2 - 1)/(x + 1) + p3) == (x - 1) + p4

    # issue 4077
    q = S('''(2*1*(x - 1/x)/(x*(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2) - 1/(x**2*(x -
        1/x)) - 2/x)) - 2*1*((x - 1/x)/((x*(x - 1/x)**2)) - 1/(x*(x -
        1/x)))*((-x + 1/x)*((x - 1/x)/((x*(x - 1/x)**2)) - 1/(x*(x -
        1/x)))/(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2) - 1/(x**2*(x - 1/x)) -
        2/x) + 1)*((x - 1/x)/((x - 1/x)**2) - ((x - 1/x)/((x*(x - 1/x)**2)) -
        1/(x*(x - 1/x)))**2/(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2) - 1/(x**2*(x
        - 1/x)) - 2/x) - 1/(x - 1/x))*(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2) -
        1/(x**2*(x - 1/x)) - 2/x)/x - 1/x)*(((-x + 1/x)/((x*(x - 1/x)**2)) +
        1/(x*(x - 1/x)))*((-(x - 1/x)/(x*(x - 1/x)) - 1/x)*((x - 1/x)/((x*(x -
        1/x)**2)) - 1/(x*(x - 1/x)))/(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2) -
        1/(x**2*(x - 1/x)) - 2/x) - 1 + (x - 1/x)/(x - 1/x))/((x*((x -
        1/x)/((x - 1/x)**2) - ((x - 1/x)/((x*(x - 1/x)**2)) - 1/(x*(x -
        1/x)))**2/(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2) - 1/(x**2*(x - 1/x)) -
        2/x) - 1/(x - 1/x))*(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2) - 1/(x**2*(x
        - 1/x)) - 2/x))) + ((x - 1/x)/((x*(x - 1/x))) + 1/x)/((x*(2*x - (-x +
        1/x)/(x**2*(x - 1/x)**2) - 1/(x**2*(x - 1/x)) - 2/x))) + 1/x)/(2*x +
        2*((x - 1/x)/((x*(x - 1/x)**2)) - 1/(x*(x - 1/x)))*((-(x - 1/x)/(x*(x
        - 1/x)) - 1/x)*((x - 1/x)/((x*(x - 1/x)**2)) - 1/(x*(x - 1/x)))/(2*x -
        (-x + 1/x)/(x**2*(x - 1/x)**2) - 1/(x**2*(x - 1/x)) - 2/x) - 1 + (x -
        1/x)/(x - 1/x))/((x*((x - 1/x)/((x - 1/x)**2) - ((x - 1/x)/((x*(x -
        1/x)**2)) - 1/(x*(x - 1/x)))**2/(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2)
        - 1/(x**2*(x - 1/x)) - 2/x) - 1/(x - 1/x))*(2*x - (-x + 1/x)/(x**2*(x
        - 1/x)**2) - 1/(x**2*(x - 1/x)) - 2/x))) - 2*((x - 1/x)/((x*(x -
        1/x))) + 1/x)/(x*(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2) - 1/(x**2*(x -
        1/x)) - 2/x)) - 2/x) - ((x - 1/x)/((x*(x - 1/x)**2)) - 1/(x*(x -
        1/x)))*((-x + 1/x)*((x - 1/x)/((x*(x - 1/x)**2)) - 1/(x*(x -
        1/x)))/(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2) - 1/(x**2*(x - 1/x)) -
        2/x) + 1)/(x*((x - 1/x)/((x - 1/x)**2) - ((x - 1/x)/((x*(x - 1/x)**2))
        - 1/(x*(x - 1/x)))**2/(2*x - (-x + 1/x)/(x**2*(x - 1/x)**2) -
        1/(x**2*(x - 1/x)) - 2/x) - 1/(x - 1/x))*(2*x - (-x + 1/x)/(x**2*(x -
        1/x)**2) - 1/(x**2*(x - 1/x)) - 2/x)) + (x - 1/x)/((x*(2*x - (-x +
        1/x)/(x**2*(x - 1/x)**2) - 1/(x**2*(x - 1/x)) - 2/x))) - 1/x''',
        evaluate=False)
    assert cancel(q, _signsimp=False) is S.NaN
    assert q.subs(x, 2) is S.NaN
    assert signsimp(q) is S.NaN

    # issue 9363
    M = MatrixSymbol('M', 5, 5)
    assert cancel(M[0,0] + 7) == M[0,0] + 7
    expr = sin(M[1, 4] + M[2, 1] * 5 * M[4, 0]) - 5 * M[1, 2] / z
    assert cancel(expr) == (z*sin(M[1, 4] + M[2, 1] * 5 * M[4, 0]) - 5 * M[1, 2]) / z

    assert cancel((x**2 + 1)/(x - I)) == x + I


def test_cancel_modulus():
    assert cancel((x**2 - 1)/(x + 1), modulus=2) == x + 1
    assert Poly(x**2 - 1, modulus=2).cancel(Poly(x + 1, modulus=2)) ==\
            (1, Poly(x + 1, modulus=2), Poly(1, x, modulus=2))


def test_make_monic_over_integers_by_scaling_roots():
    f = Poly(x**2 + 3*x + 4, x, domain='ZZ')
    g, c = f.make_monic_over_integers_by_scaling_roots()
    assert g == f
    assert c == ZZ.one

    f = Poly(x**2 + 3*x + 4, x, domain='QQ')
    g, c = f.make_monic_over_integers_by_scaling_roots()
    assert g == f.to_ring()
    assert c == ZZ.one

    f = Poly(x**2/2 + S(1)/4 * x + S(1)/8, x, domain='QQ')
    g, c = f.make_monic_over_integers_by_scaling_roots()
    assert g == Poly(x**2 + 2*x + 4, x, domain='ZZ')
    assert c == 4

    f = Poly(x**3/2 + S(1)/4 * x + S(1)/8, x, domain='QQ')
    g, c = f.make_monic_over_integers_by_scaling_roots()
    assert g == Poly(x**3 + 8*x + 16, x, domain='ZZ')
    assert c == 4

    f = Poly(x*y, x, y)
    raises(ValueError, lambda: f.make_monic_over_integers_by_scaling_roots())

    f = Poly(x, domain='RR')
    raises(ValueError, lambda: f.make_monic_over_integers_by_scaling_roots())


def test_galois_group():
    f = Poly(x ** 4 - 2)
    G, alt = f.galois_group(by_name=True)
    assert G == S4TransitiveSubgroups.D4
    assert alt is False


def test_reduced():
    f = 2*x**4 + y**2 - x**2 + y**3
    G = [x**3 - x, y**3 - y]

    Q = [2*x, 1]
    r = x**2 + y**2 + y

    assert reduced(f, G) == (Q, r)
    assert reduced(f, G, x, y) == (Q, r)

    H = groebner(G)

    assert H.reduce(f) == (Q, r)

    Q = [Poly(2*x, x, y), Poly(1, x, y)]
    r = Poly(x**2 + y**2 + y, x, y)

    assert _strict_eq(reduced(f, G, polys=True), (Q, r))
    assert _strict_eq(reduced(f, G, x, y, polys=True), (Q, r))

    H = groebner(G, polys=True)

    assert _strict_eq(H.reduce(f), (Q, r))

    f = 2*x**3 + y**3 + 3*y
    G = groebner([x**2 + y**2 - 1, x*y - 2])

    Q = [x**2 - x*y**3/2 + x*y/2 + y**6/4 - y**4/2 + y**2/4, -y**5/4 + y**3/2 + y*Rational(3, 4)]
    r = 0

    assert reduced(f, G) == (Q, r)
    assert G.reduce(f) == (Q, r)

    assert reduced(f, G, auto=False)[1] != 0
    assert G.reduce(f, auto=False)[1] != 0

    assert G.contains(f) is True
    assert G.contains(f + 1) is False

    assert reduced(1, [1], x) == ([1], 0)
    raises(ComputationFailed, lambda: reduced(1, [1]))

    f_poly = Poly(2*x**3 + y**3 + 3*y)
    G_poly = groebner([Poly(x**2 + y**2 - 1), Poly(x*y - 2)])

    Q_poly = [Poly(x**2 - 1/2*x*y**3 + 1/2*x*y + 1/4*y**6 - 1/2*y**4 + 1/4*y**2, x, y, domain='QQ'),
              Poly(-1/4*y**5 + 1/2*y**3 + 3/4*y, x, y, domain='QQ')]
    r_poly = Poly(0, x, y, domain='QQ')

    assert G_poly.reduce(f_poly) == (Q_poly, r_poly)

    Q, r = G_poly.reduce(f)
    assert all(isinstance(q, Poly) for q in Q)
    assert isinstance(r, Poly)

    f_wrong_gens = Poly(2*x**3 + y**3 + 3*y, x, y, z)
    raises(ValueError, lambda: G_poly.reduce(f_wrong_gens))

    zero_poly = Poly(0, x, y)
    Q, r = G_poly.reduce(zero_poly)
    assert all(q.is_zero for q in Q)
    assert r.is_zero

    const_poly = Poly(1, x, y)
    Q, r = G_poly.reduce(const_poly)
    assert isinstance(r, Poly)
    assert r.as_expr() == 1
    assert all(q.is_zero for q in Q)


def test_groebner():
    assert groebner([], x, y, z) == []

    assert groebner([x**2 + 1, y**4*x + x**3], x, y, order='lex') == [1 + x**2, -1 + y**4]
    assert groebner([x**2 + 1, y**4*x + x**3, x*y*z**3], x, y, z, order='grevlex') == [-1 + y**4, z**3, 1 + x**2]

    assert groebner([x**2 + 1, y**4*x + x**3], x, y, order='lex', polys=True) == \
        [Poly(1 + x**2, x, y), Poly(-1 + y**4, x, y)]
    assert groebner([x**2 + 1, y**4*x + x**3, x*y*z**3], x, y, z, order='grevlex', polys=True) == \
        [Poly(-1 + y**4, x, y, z), Poly(z**3, x, y, z), Poly(1 + x**2, x, y, z)]

    assert groebner([x**3 - 1, x**2 - 1]) == [x - 1]
    assert groebner([Eq(x**3, 1), Eq(x**2, 1)]) == [x - 1]

    F = [3*x**2 + y*z - 5*x - 1, 2*x + 3*x*y + y**2, x - 3*y + x*z - 2*z**2]
    f = z**9 - x**2*y**3 - 3*x*y**2*z + 11*y*z**2 + x**2*z**2 - 5

    G = groebner(F, x, y, z, modulus=7, symmetric=False)

    assert G == [1 + x + y + 3*z + 2*z**2 + 2*z**3 + 6*z**4 + z**5,
                 1 + 3*y + y**2 + 6*z**2 + 3*z**3 + 3*z**4 + 3*z**5 + 4*z**6,
                 1 + 4*y + 4*z + y*z + 4*z**3 + z**4 + z**6,
                 6 + 6*z + z**2 + 4*z**3 + 3*z**4 + 6*z**5 + 3*z**6 + z**7]

    Q, r = reduced(f, G, x, y, z, modulus=7, symmetric=False, polys=True)

    assert sum([ q*g for q, g in zip(Q, G.polys)], r) == Poly(f, modulus=7)

    F = [x*y - 2*y, 2*y**2 - x**2]

    assert groebner(F, x, y, order='grevlex') == \
        [y**3 - 2*y, x**2 - 2*y**2, x*y - 2*y]
    assert groebner(F, y, x, order='grevlex') == \
        [x**3 - 2*x**2, -x**2 + 2*y**2, x*y - 2*y]
    assert groebner(F, order='grevlex', field=True) == \
        [y**3 - 2*y, x**2 - 2*y**2, x*y - 2*y]

    assert groebner([1], x) == [1]

    assert groebner([x**2 + 2.0*y], x, y) == [1.0*x**2 + 2.0*y]
    raises(ComputationFailed, lambda: groebner([1]))

    assert groebner([x**2 - 1, x**3 + 1], method='buchberger') == [x + 1]
    assert groebner([x**2 - 1, x**3 + 1], method='f5b') == [x + 1]

    raises(ValueError, lambda: groebner([x, y], method='unknown'))


def test_fglm():
    F = [a + b + c + d, a*b + a*d + b*c + b*d, a*b*c + a*b*d + a*c*d + b*c*d, a*b*c*d - 1]
    G = groebner(F, a, b, c, d, order=grlex)

    B = [
        4*a + 3*d**9 - 4*d**5 - 3*d,
        4*b + 4*c - 3*d**9 + 4*d**5 + 7*d,
        4*c**2 + 3*d**10 - 4*d**6 - 3*d**2,
        4*c*d**4 + 4*c - d**9 + 4*d**5 + 5*d,
        d**12 - d**8 - d**4 + 1,
    ]

    assert groebner(F, a, b, c, d, order=lex) == B
    assert G.fglm(lex) == B

    F = [9*x**8 + 36*x**7 - 32*x**6 - 252*x**5 - 78*x**4 + 468*x**3 + 288*x**2 - 108*x + 9,
        -72*t*x**7 - 252*t*x**6 + 192*t*x**5 + 1260*t*x**4 + 312*t*x**3 - 404*t*x**2 - 576*t*x + \
        108*t - 72*x**7 - 256*x**6 + 192*x**5 + 1280*x**4 + 312*x**3 - 576*x + 96]
    G = groebner(F, t, x, order=grlex)

    B = [
        203577793572507451707*t + 627982239411707112*x**7 - 666924143779443762*x**6 - \
        10874593056632447619*x**5 + 5119998792707079562*x**4 + 72917161949456066376*x**3 + \
        20362663855832380362*x**2 - 142079311455258371571*x + 183756699868981873194,
        9*x**8 + 36*x**7 - 32*x**6 - 252*x**5 - 78*x**4 + 468*x**3 + 288*x**2 - 108*x + 9,
    ]

    assert groebner(F, t, x, order=lex) == B
    assert G.fglm(lex) == B

    F = [x**2 - x - 3*y + 1, -2*x + y**2 + y - 1]
    G = groebner(F, x, y, order=lex)

    B = [
        x**2 - x - 3*y + 1,
        y**2 - 2*x + y - 1,
    ]

    assert groebner(F, x, y, order=grlex) == B
    assert G.fglm(grlex) == B


def test_is_zero_dimensional():
    assert is_zero_dimensional([x, y], x, y) is True
    assert is_zero_dimensional([x**3 + y**2], x, y) is False

    assert is_zero_dimensional([x, y, z], x, y, z) is True
    assert is_zero_dimensional([x, y, z], x, y, z, t) is False

    F = [x*y - z, y*z - x, x*y - y]
    assert is_zero_dimensional(F, x, y, z) is True

    F = [x**2 - 2*x*z + 5, x*y**2 + y*z**3, 3*y**2 - 8*z**2]
    assert is_zero_dimensional(F, x, y, z) is True


def test_GroebnerBasis():
    F = [x*y - 2*y, 2*y**2 - x**2]

    G = groebner(F, x, y, order='grevlex')
    H = [y**3 - 2*y, x**2 - 2*y**2, x*y - 2*y]
    P = [ Poly(h, x, y) for h in H ]

    assert groebner(F + [0], x, y, order='grevlex') == G
    assert isinstance(G, GroebnerBasis) is True

    assert len(G) == 3

    assert G[0] == H[0] and not G[0].is_Poly
    assert G[1] == H[1] and not G[1].is_Poly
    assert G[2] == H[2] and not G[2].is_Poly

    assert G[1:] == H[1:] and not any(g.is_Poly for g in G[1:])
    assert G[:2] == H[:2] and not any(g.is_Poly for g in G[1:])

    assert G.exprs == H
    assert G.polys == P
    assert G.gens == (x, y)
    assert G.domain == ZZ
    assert G.order == grevlex

    assert G == H
    assert G == tuple(H)
    assert G == P
    assert G == tuple(P)

    assert G != []

    G = groebner(F, x, y, order='grevlex', polys=True)

    assert G[0] == P[0] and G[0].is_Poly
    assert G[1] == P[1] and G[1].is_Poly
    assert G[2] == P[2] and G[2].is_Poly

    assert G[1:] == P[1:] and all(g.is_Poly for g in G[1:])
    assert G[:2] == P[:2] and all(g.is_Poly for g in G[1:])


def test_poly():
    assert poly(x) == Poly(x, x)
    assert poly(y) == Poly(y, y)

    assert poly(x + y) == Poly(x + y, x, y)
    assert poly(x + sin(x)) == Poly(x + sin(x), x, sin(x))

    assert poly(x + y, wrt=y) == Poly(x + y, y, x)
    assert poly(x + sin(x), wrt=sin(x)) == Poly(x + sin(x), sin(x), x)

    assert poly(x*y + 2*x*z**2 + 17) == Poly(x*y + 2*x*z**2 + 17, x, y, z)

    assert poly(2*(y + z)**2 - 1) == Poly(2*y**2 + 4*y*z + 2*z**2 - 1, y, z)
    assert poly(
        x*(y + z)**2 - 1) == Poly(x*y**2 + 2*x*y*z + x*z**2 - 1, x, y, z)
    assert poly(2*x*(
        y + z)**2 - 1) == Poly(2*x*y**2 + 4*x*y*z + 2*x*z**2 - 1, x, y, z)

    assert poly(2*(
        y + z)**2 - x - 1) == Poly(2*y**2 + 4*y*z + 2*z**2 - x - 1, x, y, z)
    assert poly(x*(
        y + z)**2 - x - 1) == Poly(x*y**2 + 2*x*y*z + x*z**2 - x - 1, x, y, z)
    assert poly(2*x*(y + z)**2 - x - 1) == Poly(2*x*y**2 + 4*x*y*z + 2*
                x*z**2 - x - 1, x, y, z)

    assert poly(x*y + (x + y)**2 + (x + z)**2) == \
        Poly(2*x*z + 3*x*y + y**2 + z**2 + 2*x**2, x, y, z)
    assert poly(x*y*(x + y)*(x + z)**2) == \
        Poly(x**3*y**2 + x*y**2*z**2 + y*x**2*z**2 + 2*z*x**2*
             y**2 + 2*y*z*x**3 + y*x**4, x, y, z)

    assert poly(Poly(x + y + z, y, x, z)) == Poly(x + y + z, y, x, z)

    assert poly((x + y)**2, x) == Poly(x**2 + 2*x*y + y**2, x, domain=ZZ[y])
    assert poly((x + y)**2, y) == Poly(x**2 + 2*x*y + y**2, y, domain=ZZ[x])

    assert poly(1, x) == Poly(1, x)
    raises(GeneratorsNeeded, lambda: poly(1))

    # issue 6184
    assert poly(x + y, x, y) == Poly(x + y, x, y)
    assert poly(x + y, y, x) == Poly(x + y, y, x)

    # https://github.com/sympy/sympy/issues/19755
    expr1 = x + (2*x + 3)**2/5 + S(6)/5
    assert poly(expr1).as_expr() == expr1.expand()
    expr2 = y*(y+1) + S(1)/3
    assert poly(expr2).as_expr() == expr2.expand()


def test_keep_coeff():
    u = Mul(2, x + 1, evaluate=False)
    assert _keep_coeff(S.One, x) == x
    assert _keep_coeff(S.NegativeOne, x) == -x
    assert _keep_coeff(S(1.0), x) == 1.0*x
    assert _keep_coeff(S(-1.0), x) == -1.0*x
    assert _keep_coeff(S.One, 2*x) == 2*x
    assert _keep_coeff(S(2), x/2) == x
    assert _keep_coeff(S(2), sin(x)) == 2*sin(x)
    assert _keep_coeff(S(2), x + 1) == u
    assert _keep_coeff(x, 1/x) == 1
    assert _keep_coeff(x + 1, S(2)) == u
    assert _keep_coeff(S.Half, S.One) == S.Half
    p = Pow(2, 3, evaluate=False)
    assert _keep_coeff(S(-1), p) == Mul(-1, p, evaluate=False)
    a = Add(2, p, evaluate=False)
    assert _keep_coeff(S.Half, a, clear=True
        ) == Mul(S.Half, a, evaluate=False)
    assert _keep_coeff(S.Half, a, clear=False
        ) == Add(1, Mul(S.Half, p, evaluate=False), evaluate=False)


def test_poly_matching_consistency():
    # Test for this issue:
    # https://github.com/sympy/sympy/issues/5514
    assert I * Poly(x, x) == Poly(I*x, x)
    assert Poly(x, x) * I == Poly(I*x, x)


def test_issue_5786():
    assert expand(factor(expand(
        (x - I*y)*(z - I*t)), extension=[I])) == -I*t*x - t*y + x*z - I*y*z


def test_noncommutative():
    class foo(Expr):
        is_commutative=False
    e = x/(x + x*y)
    c = 1/( 1 + y)
    assert cancel(foo(e)) == foo(c)
    assert cancel(e + foo(e)) == c + foo(c)
    assert cancel(e*foo(c)) == c*foo(c)


def test_to_rational_coeffs():
    assert to_rational_coeffs(
        Poly(x**3 + y*x**2 + sqrt(y), x, domain='EX')) is None
    # issue 21268
    assert to_rational_coeffs(
        Poly(y**3 + sqrt(2)*y**2*sin(x) + 1, y)) is None

    assert to_rational_coeffs(Poly(x, y)) is None
    assert to_rational_coeffs(Poly(sqrt(2)*y)) is None


def test_factor_terms():
    # issue 7067
    assert factor_list(x*(x + y)) == (1, [(x, 1), (x + y, 1)])
    assert sqf_list(x*(x + y)) == (1, [(x**2 + x*y, 1)])


def test_as_list():
    # issue 14496
    assert Poly(x**3 + 2, x, domain='ZZ').as_list() == [1, 0, 0, 2]
    assert Poly(x**2 + y + 1, x, y, domain='ZZ').as_list() == [[1], [], [1, 1]]
    assert Poly(x**2 + y + 1, x, y, z, domain='ZZ').as_list() == \
                                                    [[[1]], [[]], [[1], [1]]]


def test_issue_11198():
    assert factor_list(sqrt(2)*x) == (sqrt(2), [(x, 1)])
    assert factor_list(sqrt(2)*sin(x), sin(x)) == (sqrt(2), [(sin(x), 1)])


def test_Poly_precision():
    # Make sure Poly doesn't lose precision
    p = Poly(pi.evalf(100)*x)
    assert p.as_expr() == pi.evalf(100)*x


def test_issue_12400():
    # Correction of check for negative exponents
    assert poly(1/(1+sqrt(2)), x) == \
            Poly(1/(1+sqrt(2)), x, domain='EX')

def test_issue_14364():
    assert gcd(S(6)*(1 + sqrt(3))/5, S(3)*(1 + sqrt(3))/10) == Rational(3, 10) * (1 + sqrt(3))
    assert gcd(sqrt(5)*Rational(4, 7), sqrt(5)*Rational(2, 3)) == sqrt(5)*Rational(2, 21)

    assert lcm(Rational(2, 3)*sqrt(3), Rational(5, 6)*sqrt(3)) == S(10)*sqrt(3)/3
    assert lcm(3*sqrt(3), 4/sqrt(3)) == 12*sqrt(3)
    assert lcm(S(5)*(1 + 2**Rational(1, 3))/6, S(3)*(1 + 2**Rational(1, 3))/8) == Rational(15, 2) * (1 + 2**Rational(1, 3))

    assert gcd(Rational(2, 3)*sqrt(3), Rational(5, 6)/sqrt(3)) == sqrt(3)/18
    assert gcd(S(4)*sqrt(13)/7, S(3)*sqrt(13)/14) == sqrt(13)/14

    # gcd_list and lcm_list
    assert gcd([S(2)*sqrt(47)/7, S(6)*sqrt(47)/5, S(8)*sqrt(47)/5]) == sqrt(47)*Rational(2, 35)
    assert gcd([S(6)*(1 + sqrt(7))/5, S(2)*(1 + sqrt(7))/7, S(4)*(1 + sqrt(7))/13]) ==  (1 + sqrt(7))*Rational(2, 455)
    assert lcm((Rational(7, 2)/sqrt(15), Rational(5, 6)/sqrt(15), Rational(5, 8)/sqrt(15))) == Rational(35, 2)/sqrt(15)
    assert lcm([S(5)*(2 + 2**Rational(5, 7))/6, S(7)*(2 + 2**Rational(5, 7))/2, S(13)*(2 + 2**Rational(5, 7))/4]) == Rational(455, 2) * (2 + 2**Rational(5, 7))


def test_issue_15669():
    x = Symbol("x", positive=True)
    expr = (16*x**3/(-x**2 + sqrt(8*x**2 + (x**2 - 2)**2) + 2)**2 -
        2*2**Rational(4, 5)*x*(-x**2 + sqrt(8*x**2 + (x**2 - 2)**2) + 2)**Rational(3, 5) + 10*x)
    assert factor(expr, deep=True) == x*(x**2 + 2)


def test_issue_17988():
    x = Symbol('x')
    p = poly(x - 1)
    with warns_deprecated_sympy():
        M = Matrix([[poly(x + 1), poly(x + 1)]])
    with warns(SymPyDeprecationWarning, test_stacklevel=False):
        assert p * M == M * p == Matrix([[poly(x**2 - 1), poly(x**2 - 1)]])


def test_issue_18205():
    assert cancel((2 + I)*(3 - I)) == 7 + I
    assert cancel((2 + I)*(2 - I)) == 5


def test_issue_8695():
    p = (x**2 + 1) * (x - 1)**2 * (x - 2)**3 * (x - 3)**3
    result = (1, [(x**2 + 1, 1), (x - 1, 2), (x**2 - 5*x + 6, 3)])
    assert sqf_list(p) == result


def test_issue_19113():
    eq = sin(x)**3 - sin(x) + 1
    raises(PolynomialError, lambda: refine_root(eq, 1, 2, 1e-2))
    raises(PolynomialError, lambda: count_roots(eq, -1, 1))
    raises(PolynomialError, lambda: real_roots(eq))
    raises(PolynomialError, lambda: nroots(eq))
    raises(PolynomialError, lambda: ground_roots(eq))
    raises(PolynomialError, lambda: nth_power_roots_poly(eq, 2))


def test_issue_19360():
    f = 2*x**2 - 2*sqrt(2)*x*y + y**2
    assert factor(f, extension=sqrt(2)) == 2*(x - (sqrt(2)*y/2))**2

    f = -I*t*x - t*y + x*z - I*y*z
    assert factor(f, extension=I) == (x - I*y)*(-I*t + z)


def test_poly_copy_equals_original():
    poly = Poly(x + y, x, y, z)
    copy = poly.copy()
    assert poly == copy, (
        "Copied polynomial not equal to original.")
    assert poly.gens == copy.gens, (
        "Copied polynomial has different generators than original.")


def test_deserialized_poly_equals_original():
    poly = Poly(x + y, x, y, z)
    deserialized = pickle.loads(pickle.dumps(poly))
    assert poly == deserialized, (
        "Deserialized polynomial not equal to original.")
    assert poly.gens == deserialized.gens, (
        "Deserialized polynomial has different generators than original.")


def test_issue_20389():
    result = degree(x * (x + 1) - x ** 2 - x, x)
    assert result == -oo


def test_issue_20985():
    from sympy.core.symbol import symbols
    w, R = symbols('w R')
    poly = Poly(1.0 + I*w/R, w, 1/R)
    assert poly.degree() == S(1)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\click\__init__.py ===
"""
Click is a simple Python module inspired by the stdlib optparse to make
writing command line scripts fun. Unlike other modules, it's based
around a simple API that does not come with too much magic and is
composable.
"""

from __future__ import annotations

from .core import Argument as Argument
from .core import Command as Command
from .core import CommandCollection as CommandCollection
from .core import Context as Context
from .core import Group as Group
from .core import Option as Option
from .core import Parameter as Parameter
from .decorators import argument as argument
from .decorators import command as command
from .decorators import confirmation_option as confirmation_option
from .decorators import group as group
from .decorators import help_option as help_option
from .decorators import make_pass_decorator as make_pass_decorator
from .decorators import option as option
from .decorators import pass_context as pass_context
from .decorators import pass_obj as pass_obj
from .decorators import password_option as password_option
from .decorators import version_option as version_option
from .exceptions import Abort as Abort
from .exceptions import BadArgumentUsage as BadArgumentUsage
from .exceptions import BadOptionUsage as BadOptionUsage
from .exceptions import BadParameter as BadParameter
from .exceptions import ClickException as ClickException
from .exceptions import FileError as FileError
from .exceptions import MissingParameter as MissingParameter
from .exceptions import NoSuchOption as NoSuchOption
from .exceptions import UsageError as UsageError
from .formatting import HelpFormatter as HelpFormatter
from .formatting import wrap_text as wrap_text
from .globals import get_current_context as get_current_context
from .termui import clear as clear
from .termui import confirm as confirm
from .termui import echo_via_pager as echo_via_pager
from .termui import edit as edit
from .termui import getchar as getchar
from .termui import launch as launch
from .termui import pause as pause
from .termui import progressbar as progressbar
from .termui import prompt as prompt
from .termui import secho as secho
from .termui import style as style
from .termui import unstyle as unstyle
from .types import BOOL as BOOL
from .types import Choice as Choice
from .types import DateTime as DateTime
from .types import File as File
from .types import FLOAT as FLOAT
from .types import FloatRange as FloatRange
from .types import INT as INT
from .types import IntRange as IntRange
from .types import ParamType as ParamType
from .types import Path as Path
from .types import STRING as STRING
from .types import Tuple as Tuple
from .types import UNPROCESSED as UNPROCESSED
from .types import UUID as UUID
from .utils import echo as echo
from .utils import format_filename as format_filename
from .utils import get_app_dir as get_app_dir
from .utils import get_binary_stream as get_binary_stream
from .utils import get_text_stream as get_text_stream
from .utils import open_file as open_file


def __getattr__(name: str) -> object:
    import warnings

    if name == "BaseCommand":
        from .core import _BaseCommand

        warnings.warn(
            "'BaseCommand' is deprecated and will be removed in Click 9.0. Use"
            " 'Command' instead.",
            DeprecationWarning,
            stacklevel=2,
        )
        return _BaseCommand

    if name == "MultiCommand":
        from .core import _MultiCommand

        warnings.warn(
            "'MultiCommand' is deprecated and will be removed in Click 9.0. Use"
            " 'Group' instead.",
            DeprecationWarning,
            stacklevel=2,
        )
        return _MultiCommand

    if name == "OptionParser":
        from .parser import _OptionParser

        warnings.warn(
            "'OptionParser' is deprecated and will be removed in Click 9.0. The"
            " old parser is available in 'optparse'.",
            DeprecationWarning,
            stacklevel=2,
        )
        return _OptionParser

    if name == "__version__":
        import importlib.metadata
        import warnings

        warnings.warn(
            "The '__version__' attribute is deprecated and will be removed in"
            " Click 9.1. Use feature detection or"
            " 'importlib.metadata.version(\"click\")' instead.",
            DeprecationWarning,
            stacklevel=2,
        )
        return importlib.metadata.version("click")

    raise AttributeError(name)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\tensor\array\expressions\utils.py ===
import bisect
from collections import defaultdict

from sympy.combinatorics import Permutation
from sympy.core.containers import Tuple
from sympy.core.numbers import Integer


def _get_mapping_from_subranks(subranks):
    mapping = {}
    counter = 0
    for i, rank in enumerate(subranks):
        for j in range(rank):
            mapping[counter] = (i, j)
            counter += 1
    return mapping


def _get_contraction_links(args, subranks, *contraction_indices):
    mapping = _get_mapping_from_subranks(subranks)
    contraction_tuples = [[mapping[j] for j in i] for i in contraction_indices]
    dlinks = defaultdict(dict)
    for links in contraction_tuples:
        if len(links) == 2:
            (arg1, pos1), (arg2, pos2) = links
            dlinks[arg1][pos1] = (arg2, pos2)
            dlinks[arg2][pos2] = (arg1, pos1)
            continue

    return args, dict(dlinks)


def _sort_contraction_indices(pairing_indices):
    pairing_indices = [Tuple(*sorted(i)) for i in pairing_indices]
    pairing_indices.sort(key=lambda x: min(x))
    return pairing_indices


def _get_diagonal_indices(flattened_indices):
    axes_contraction = defaultdict(list)
    for i, ind in enumerate(flattened_indices):
        if isinstance(ind, (int, Integer)):
            # If the indices is a number, there can be no diagonal operation:
            continue
        axes_contraction[ind].append(i)
    axes_contraction = {k: v for k, v in axes_contraction.items() if len(v) > 1}
    # Put the diagonalized indices at the end:
    ret_indices = [i for i in flattened_indices if i not in axes_contraction]
    diag_indices = list(axes_contraction)
    diag_indices.sort(key=lambda x: flattened_indices.index(x))
    diagonal_indices = [tuple(axes_contraction[i]) for i in diag_indices]
    ret_indices += diag_indices
    ret_indices = tuple(ret_indices)
    return diagonal_indices, ret_indices


def _get_argindex(subindices, ind):
    for i, sind in enumerate(subindices):
        if ind == sind:
            return i
        if isinstance(sind, (set, frozenset)) and ind in sind:
            return i
    raise IndexError("%s not found in %s" % (ind, subindices))


def _apply_recursively_over_nested_lists(func, arr):
    if isinstance(arr, (tuple, list, Tuple)):
        return tuple(_apply_recursively_over_nested_lists(func, i) for i in arr)
    elif isinstance(arr, Tuple):
        return Tuple.fromiter(_apply_recursively_over_nested_lists(func, i) for i in arr)
    else:
        return func(arr)


def _build_push_indices_up_func_transformation(flattened_contraction_indices):
    shifts = {0: 0}
    i = 0
    cumulative = 0
    while i < len(flattened_contraction_indices):
        j = 1
        while i+j < len(flattened_contraction_indices):
            if flattened_contraction_indices[i] + j != flattened_contraction_indices[i+j]:
                break
            j += 1
        cumulative += j
        shifts[flattened_contraction_indices[i]] = cumulative
        i += j
    shift_keys = sorted(shifts.keys())

    def func(idx):
        return shifts[shift_keys[bisect.bisect_right(shift_keys, idx)-1]]

    def transform(j):
        if j in flattened_contraction_indices:
            return None
        else:
            return j - func(j)

    return transform


def _build_push_indices_down_func_transformation(flattened_contraction_indices):
    N = flattened_contraction_indices[-1]+2

    shifts = [i for i in range(N) if i not in flattened_contraction_indices]

    def transform(j):
        if j < len(shifts):
            return shifts[j]
        else:
            return j + shifts[-1] - len(shifts) + 1

    return transform


def _apply_permutation_to_list(perm: Permutation, target_list: list):
    """
    Permute a list according to the given permutation.
    """
    new_list = [None for i in range(perm.size)]
    for i, e in enumerate(target_list):
        new_list[perm(i)] = e
    return new_list

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_subprocess_platform\__init__.py ===
# Platform-specific subprocess bits'n'pieces.
from __future__ import annotations

import os
import sys
from typing import TYPE_CHECKING

import trio

from .. import _core, _subprocess
from .._abc import ReceiveStream, SendStream  # noqa: TC001

_wait_child_exiting_error: ImportError | None = None
_create_child_pipe_error: ImportError | None = None


if TYPE_CHECKING:
    # internal types for the pipe representations used in type checking only
    class ClosableSendStream(SendStream):
        def close(self) -> None: ...

    class ClosableReceiveStream(ReceiveStream):
        def close(self) -> None: ...


# Fallback versions of the functions provided -- implementations
# per OS are imported atop these at the bottom of the module.
async def wait_child_exiting(process: _subprocess.Process) -> None:
    """Block until the child process managed by ``process`` is exiting.

    It is invalid to call this function if the process has already
    been waited on; that is, ``process.returncode`` must be None.

    When this function returns, it indicates that a call to
    :meth:`subprocess.Popen.wait` will immediately be able to
    return the process's exit status. The actual exit status is not
    consumed by this call, since :class:`~subprocess.Popen` wants
    to be able to do that itself.
    """
    raise NotImplementedError from _wait_child_exiting_error  # pragma: no cover


def create_pipe_to_child_stdin() -> tuple[ClosableSendStream, int]:
    """Create a new pipe suitable for sending data from this
    process to the standard input of a child we're about to spawn.

    Returns:
      A pair ``(trio_end, subprocess_end)`` where ``trio_end`` is a
      :class:`~trio.abc.SendStream` and ``subprocess_end`` is
      something suitable for passing as the ``stdin`` argument of
      :class:`subprocess.Popen`.
    """
    raise NotImplementedError from _create_child_pipe_error  # pragma: no cover


def create_pipe_from_child_output() -> tuple[ClosableReceiveStream, int]:
    """Create a new pipe suitable for receiving data into this
    process from the standard output or error stream of a child
    we're about to spawn.

    Returns:
      A pair ``(trio_end, subprocess_end)`` where ``trio_end`` is a
      :class:`~trio.abc.ReceiveStream` and ``subprocess_end`` is
      something suitable for passing as the ``stdin`` argument of
      :class:`subprocess.Popen`.
    """
    raise NotImplementedError from _create_child_pipe_error  # pragma: no cover


try:
    if sys.platform == "win32":
        from .windows import wait_child_exiting
    elif sys.platform != "linux" and (TYPE_CHECKING or hasattr(_core, "wait_kevent")):
        from .kqueue import wait_child_exiting
    else:
        # as it's an exported symbol, noqa'd
        from .waitid import wait_child_exiting  # noqa: F401
except ImportError as ex:  # pragma: no cover
    _wait_child_exiting_error = ex

try:
    if TYPE_CHECKING:
        # Not worth type checking these definitions
        pass

    elif os.name == "posix":

        def create_pipe_to_child_stdin() -> tuple[trio.lowlevel.FdStream, int]:
            rfd, wfd = os.pipe()
            return trio.lowlevel.FdStream(wfd), rfd

        def create_pipe_from_child_output() -> tuple[trio.lowlevel.FdStream, int]:
            rfd, wfd = os.pipe()
            return trio.lowlevel.FdStream(rfd), wfd

    elif os.name == "nt":
        import msvcrt

        # This isn't exported or documented, but it's also not
        # underscore-prefixed, and seems kosher to use. The asyncio docs
        # for 3.5 included an example that imported socketpair from
        # windows_utils (before socket.socketpair existed on Windows), and
        # when asyncio.windows_utils.socketpair was removed in 3.7, the
        # removal was mentioned in the release notes.
        from asyncio.windows_utils import pipe as windows_pipe

        from .._windows_pipes import PipeReceiveStream, PipeSendStream

        def create_pipe_to_child_stdin() -> tuple[PipeSendStream, int]:
            # for stdin, we want the write end (our end) to use overlapped I/O
            rh, wh = windows_pipe(overlapped=(False, True))
            return PipeSendStream(wh), msvcrt.open_osfhandle(rh, os.O_RDONLY)

        def create_pipe_from_child_output() -> tuple[PipeReceiveStream, int]:
            # for stdout/err, it's the read end that's overlapped
            rh, wh = windows_pipe(overlapped=(True, False))
            return PipeReceiveStream(rh), msvcrt.open_osfhandle(wh, 0)

    else:  # pragma: no cover
        raise ImportError("pipes not implemented on this platform")

except ImportError as ex:  # pragma: no cover
    _create_child_pipe_error = ex

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\idna\codec.py ===
import codecs
import re
from typing import Any, Optional, Tuple

from .core import IDNAError, alabel, decode, encode, ulabel

_unicode_dots_re = re.compile("[\u002e\u3002\uff0e\uff61]")


class Codec(codecs.Codec):
    def encode(self, data: str, errors: str = "strict") -> Tuple[bytes, int]:
        if errors != "strict":
            raise IDNAError('Unsupported error handling "{}"'.format(errors))

        if not data:
            return b"", 0

        return encode(data), len(data)

    def decode(self, data: bytes, errors: str = "strict") -> Tuple[str, int]:
        if errors != "strict":
            raise IDNAError('Unsupported error handling "{}"'.format(errors))

        if not data:
            return "", 0

        return decode(data), len(data)


class IncrementalEncoder(codecs.BufferedIncrementalEncoder):
    def _buffer_encode(self, data: str, errors: str, final: bool) -> Tuple[bytes, int]:
        if errors != "strict":
            raise IDNAError('Unsupported error handling "{}"'.format(errors))

        if not data:
            return b"", 0

        labels = _unicode_dots_re.split(data)
        trailing_dot = b""
        if labels:
            if not labels[-1]:
                trailing_dot = b"."
                del labels[-1]
            elif not final:
                # Keep potentially unfinished label until the next call
                del labels[-1]
                if labels:
                    trailing_dot = b"."

        result = []
        size = 0
        for label in labels:
            result.append(alabel(label))
            if size:
                size += 1
            size += len(label)

        # Join with U+002E
        result_bytes = b".".join(result) + trailing_dot
        size += len(trailing_dot)
        return result_bytes, size


class IncrementalDecoder(codecs.BufferedIncrementalDecoder):
    def _buffer_decode(self, data: Any, errors: str, final: bool) -> Tuple[str, int]:
        if errors != "strict":
            raise IDNAError('Unsupported error handling "{}"'.format(errors))

        if not data:
            return ("", 0)

        if not isinstance(data, str):
            data = str(data, "ascii")

        labels = _unicode_dots_re.split(data)
        trailing_dot = ""
        if labels:
            if not labels[-1]:
                trailing_dot = "."
                del labels[-1]
            elif not final:
                # Keep potentially unfinished label until the next call
                del labels[-1]
                if labels:
                    trailing_dot = "."

        result = []
        size = 0
        for label in labels:
            result.append(ulabel(label))
            if size:
                size += 1
            size += len(label)

        result_str = ".".join(result) + trailing_dot
        size += len(trailing_dot)
        return (result_str, size)


class StreamWriter(Codec, codecs.StreamWriter):
    pass


class StreamReader(Codec, codecs.StreamReader):
    pass


def search_function(name: str) -> Optional[codecs.CodecInfo]:
    if name != "idna2008":
        return None
    return codecs.CodecInfo(
        name=name,
        encode=Codec().encode,
        decode=Codec().decode,
        incrementalencoder=IncrementalEncoder,
        incrementaldecoder=IncrementalDecoder,
        streamwriter=StreamWriter,
        streamreader=StreamReader,
    )


codecs.register(search_function)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_qordered_layernorm.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
from logging import getLogger

from fusion_base import Fusion
from fusion_utils import FusionUtils
from onnx import helper
from onnx_model import OnnxModel

logger = getLogger(__name__)


class FusionQOrderedLayerNormalization(Fusion):
    def __init__(self, model: OnnxModel):
        super().__init__(model, "QOrderedLayerNormalization", "LayerNormalization")

    def fuse(self, node, input_name_to_nodes: dict, output_name_to_node: dict):
        """
        Fuse (quantized) Layer Normalization subgraph into one node QOrderedLayerNormalization:
            quantized input  -> DQ
                                |
                                |
            (other inputs)-> LayerNormalization --> Q -->

            should become

            (quantized input + other inputs)->  QOrderedLayerNormalization --> Q -->
        """

        children = self.model.get_children(node, input_name_to_nodes)

        # Should only have 1 child - QuantizeLinear (or)
        # Should have 2 children - QuantizeLinear + Shape
        if not (
            (len(children) == 1 and children[0].op_type == "QuantizeLinear")
            or (len(children) == 2 and children[0].op_type == "QuantizeLinear" and children[1].op_type == "Shape")
        ):
            return

        downstream_quantize_node = children[0]
        downstream_shape_node = None

        if len(children) == 2:
            downstream_shape_node = children[1]

        if not FusionUtils.check_qdq_node_for_fusion(downstream_quantize_node, self.model):
            return

        # The first input to LayerNormalization should flow through a DequantizeLinear node
        first_path_id, first_input_parent_nodes, _ = self.model.match_parent_paths(
            node,
            [(["DequantizeLinear"], [0])],
            output_name_to_node,
        )

        if first_path_id < 0:
            return

        upstream_dequantize_node = first_input_parent_nodes[0]

        if not FusionUtils.check_qdq_node_for_fusion(upstream_dequantize_node, self.model):
            return

        # Fusion logic
        subgraph_nodes = [node]  # LayerNormalization
        subgraph_nodes.extend([downstream_quantize_node])  # Q node after LayerNormalization

        upstream_dequantize_node_children = self.model.get_children(upstream_dequantize_node, input_name_to_nodes)

        # In GPT2, the DQ node will be feeding a residual downstream Add and hence,
        # we do not want to remove it
        if len(upstream_dequantize_node_children) == 1:
            subgraph_nodes.extend([upstream_dequantize_node])  # DQ node before LayerNormalization

        if not self.model.is_safe_to_fuse_nodes(
            subgraph_nodes,
            (
                [node.output[0], downstream_quantize_node.output[0]]
                if downstream_shape_node is not None
                else downstream_quantize_node.output
            ),
            input_name_to_nodes,
            output_name_to_node,
        ):
            logger.debug("It is not safe to fuse QOrderedLayerNormalization node. Skip")
            return

        self.nodes_to_remove.extend(subgraph_nodes)

        normalize_node = helper.make_node(
            "QOrderedLayerNormalization",
            inputs=[
                upstream_dequantize_node.input[0],
                upstream_dequantize_node.input[1],
                node.input[1],
                node.input[2],
                downstream_quantize_node.input[1],
            ],
            outputs=[downstream_quantize_node.output[0]],
            name=self.model.create_node_name("QOrderedLayerNormalization", name_prefix="QOrderedLayerNormalization"),
        )

        # Arrange the downstream Shape's input to be fed from the
        # downstream QuantizeLinear node, so that fusion will
        # be deemed safe
        if downstream_shape_node is not None:
            self.model.replace_node_input(
                downstream_shape_node, downstream_shape_node.input[0], downstream_quantize_node.output[0]
            )

        # TODO: We only support CuBlasLt order ORDER_ROW for now.
        # Once we start supporting other data ordering format(s), we
        # will support user configuring the data ordering for the op.
        normalize_node.attribute.extend([helper.make_attribute("order_X", 1)])
        normalize_node.attribute.extend([helper.make_attribute("order_Y", 1)])

        normalize_node.domain = "com.microsoft"

        self.nodes_to_add.append(normalize_node)
        self.node_name_to_graph_name[normalize_node.name] = self.this_graph_name

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\idna\codec.py ===
import codecs
import re
from typing import Any, Optional, Tuple

from .core import IDNAError, alabel, decode, encode, ulabel

_unicode_dots_re = re.compile("[\u002e\u3002\uff0e\uff61]")


class Codec(codecs.Codec):
    def encode(self, data: str, errors: str = "strict") -> Tuple[bytes, int]:
        if errors != "strict":
            raise IDNAError('Unsupported error handling "{}"'.format(errors))

        if not data:
            return b"", 0

        return encode(data), len(data)

    def decode(self, data: bytes, errors: str = "strict") -> Tuple[str, int]:
        if errors != "strict":
            raise IDNAError('Unsupported error handling "{}"'.format(errors))

        if not data:
            return "", 0

        return decode(data), len(data)


class IncrementalEncoder(codecs.BufferedIncrementalEncoder):
    def _buffer_encode(self, data: str, errors: str, final: bool) -> Tuple[bytes, int]:
        if errors != "strict":
            raise IDNAError('Unsupported error handling "{}"'.format(errors))

        if not data:
            return b"", 0

        labels = _unicode_dots_re.split(data)
        trailing_dot = b""
        if labels:
            if not labels[-1]:
                trailing_dot = b"."
                del labels[-1]
            elif not final:
                # Keep potentially unfinished label until the next call
                del labels[-1]
                if labels:
                    trailing_dot = b"."

        result = []
        size = 0
        for label in labels:
            result.append(alabel(label))
            if size:
                size += 1
            size += len(label)

        # Join with U+002E
        result_bytes = b".".join(result) + trailing_dot
        size += len(trailing_dot)
        return result_bytes, size


class IncrementalDecoder(codecs.BufferedIncrementalDecoder):
    def _buffer_decode(self, data: Any, errors: str, final: bool) -> Tuple[str, int]:
        if errors != "strict":
            raise IDNAError('Unsupported error handling "{}"'.format(errors))

        if not data:
            return ("", 0)

        if not isinstance(data, str):
            data = str(data, "ascii")

        labels = _unicode_dots_re.split(data)
        trailing_dot = ""
        if labels:
            if not labels[-1]:
                trailing_dot = "."
                del labels[-1]
            elif not final:
                # Keep potentially unfinished label until the next call
                del labels[-1]
                if labels:
                    trailing_dot = "."

        result = []
        size = 0
        for label in labels:
            result.append(ulabel(label))
            if size:
                size += 1
            size += len(label)

        result_str = ".".join(result) + trailing_dot
        size += len(trailing_dot)
        return (result_str, size)


class StreamWriter(Codec, codecs.StreamWriter):
    pass


class StreamReader(Codec, codecs.StreamReader):
    pass


def search_function(name: str) -> Optional[codecs.CodecInfo]:
    if name != "idna2008":
        return None
    return codecs.CodecInfo(
        name=name,
        encode=Codec().encode,
        decode=Codec().decode,
        incrementalencoder=IncrementalEncoder,
        incrementaldecoder=IncrementalDecoder,
        streamwriter=StreamWriter,
        streamreader=StreamReader,
    )


codecs.register(search_function)