
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\integrals\rationaltools.py ===
"""This module implements tools for integrating rational functions. """

from sympy.core.function import Lambda
from sympy.core.numbers import I
from sympy.core.singleton import S
from sympy.core.symbol import (Dummy, Symbol, symbols)
from sympy.functions.elementary.exponential import log
from sympy.functions.elementary.trigonometric import atan
from sympy.polys.polyerrors import DomainError
from sympy.polys.polyroots import roots
from sympy.polys.polytools import cancel
from sympy.polys.rootoftools import RootSum
from sympy.polys import Poly, resultant, ZZ


def ratint(f, x, **flags):
    """
    Performs indefinite integration of rational functions.

    Explanation
    ===========

    Given a field :math:`K` and a rational function :math:`f = p/q`,
    where :math:`p` and :math:`q` are polynomials in :math:`K[x]`,
    returns a function :math:`g` such that :math:`f = g'`.

    Examples
    ========

    >>> from sympy.integrals.rationaltools import ratint
    >>> from sympy.abc import x

    >>> ratint(36/(x**5 - 2*x**4 - 2*x**3 + 4*x**2 + x - 2), x)
    (12*x + 6)/(x**2 - 1) + 4*log(x - 2) - 4*log(x + 1)

    References
    ==========

    .. [1] M. Bronstein, Symbolic Integration I: Transcendental
       Functions, Second Edition, Springer-Verlag, 2005, pp. 35-70

    See Also
    ========

    sympy.integrals.integrals.Integral.doit
    sympy.integrals.rationaltools.ratint_logpart
    sympy.integrals.rationaltools.ratint_ratpart

    """
    if isinstance(f, tuple):
        p, q = f
    else:
        p, q = f.as_numer_denom()

    p, q = Poly(p, x, composite=False, field=True), Poly(q, x, composite=False, field=True)

    coeff, p, q = p.cancel(q)
    poly, p = p.div(q)

    result = poly.integrate(x).as_expr()

    if p.is_zero:
        return coeff*result

    g, h = ratint_ratpart(p, q, x)

    P, Q = h.as_numer_denom()

    P = Poly(P, x)
    Q = Poly(Q, x)

    q, r = P.div(Q)

    result += g + q.integrate(x).as_expr()

    if not r.is_zero:
        symbol = flags.get('symbol', 't')

        if not isinstance(symbol, Symbol):
            t = Dummy(symbol)
        else:
            t = symbol.as_dummy()

        L = ratint_logpart(r, Q, x, t)

        real = flags.get('real')

        if real is None:
            if isinstance(f, tuple):
                p, q = f
                atoms = p.atoms() | q.atoms()
            else:
                atoms = f.atoms()

            for elt in atoms - {x}:
                if not elt.is_extended_real:
                    real = False
                    break
            else:
                real = True

        eps = S.Zero

        if not real:
            for h, q in L:
                _, h = h.primitive()
                eps += RootSum(
                    q, Lambda(t, t*log(h.as_expr())), quadratic=True)
        else:
            for h, q in L:
                _, h = h.primitive()
                R = log_to_real(h, q, x, t)

                if R is not None:
                    eps += R
                else:
                    eps += RootSum(
                        q, Lambda(t, t*log(h.as_expr())), quadratic=True)

        result += eps

    return coeff*result


def ratint_ratpart(f, g, x):
    """
    Horowitz-Ostrogradsky algorithm.

    Explanation
    ===========

    Given a field K and polynomials f and g in K[x], such that f and g
    are coprime and deg(f) < deg(g), returns fractions A and B in K(x),
    such that f/g = A' + B and B has square-free denominator.

    Examples
    ========

        >>> from sympy.integrals.rationaltools import ratint_ratpart
        >>> from sympy.abc import x, y
        >>> from sympy import Poly
        >>> ratint_ratpart(Poly(1, x, domain='ZZ'),
        ... Poly(x + 1, x, domain='ZZ'), x)
        (0, 1/(x + 1))
        >>> ratint_ratpart(Poly(1, x, domain='EX'),
        ... Poly(x**2 + y**2, x, domain='EX'), x)
        (0, 1/(x**2 + y**2))
        >>> ratint_ratpart(Poly(36, x, domain='ZZ'),
        ... Poly(x**5 - 2*x**4 - 2*x**3 + 4*x**2 + x - 2, x, domain='ZZ'), x)
        ((12*x + 6)/(x**2 - 1), 12/(x**2 - x - 2))

    See Also
    ========

    ratint, ratint_logpart
    """
    from sympy.solvers.solvers import solve

    f = Poly(f, x)
    g = Poly(g, x)

    u, v, _ = g.cofactors(g.diff())

    n = u.degree()
    m = v.degree()

    A_coeffs = [ Dummy('a' + str(n - i)) for i in range(0, n) ]
    B_coeffs = [ Dummy('b' + str(m - i)) for i in range(0, m) ]

    C_coeffs = A_coeffs + B_coeffs

    A = Poly(A_coeffs, x, domain=ZZ[C_coeffs])
    B = Poly(B_coeffs, x, domain=ZZ[C_coeffs])

    H = f - A.diff()*v + A*(u.diff()*v).quo(u) - B*u

    result = solve(H.coeffs(), C_coeffs)

    A = A.as_expr().subs(result)
    B = B.as_expr().subs(result)

    rat_part = cancel(A/u.as_expr(), x)
    log_part = cancel(B/v.as_expr(), x)

    return rat_part, log_part


def ratint_logpart(f, g, x, t=None):
    r"""
    Lazard-Rioboo-Trager algorithm.

    Explanation
    ===========

    Given a field K and polynomials f and g in K[x], such that f and g
    are coprime, deg(f) < deg(g) and g is square-free, returns a list
    of tuples (s_i, q_i) of polynomials, for i = 1..n, such that s_i
    in K[t, x] and q_i in K[t], and::

                           ___    ___
                 d  f   d  \  `   \  `
                 -- - = --  )      )   a log(s_i(a, x))
                 dx g   dx /__,   /__,
                          i=1..n a | q_i(a) = 0

    Examples
    ========

    >>> from sympy.integrals.rationaltools import ratint_logpart
    >>> from sympy.abc import x
    >>> from sympy import Poly
    >>> ratint_logpart(Poly(1, x, domain='ZZ'),
    ... Poly(x**2 + x + 1, x, domain='ZZ'), x)
    [(Poly(x + 3*_t/2 + 1/2, x, domain='QQ[_t]'),
    ...Poly(3*_t**2 + 1, _t, domain='ZZ'))]
    >>> ratint_logpart(Poly(12, x, domain='ZZ'),
    ... Poly(x**2 - x - 2, x, domain='ZZ'), x)
    [(Poly(x - 3*_t/8 - 1/2, x, domain='QQ[_t]'),
    ...Poly(-_t**2 + 16, _t, domain='ZZ'))]

    See Also
    ========

    ratint, ratint_ratpart
    """
    f, g = Poly(f, x), Poly(g, x)

    t = t or Dummy('t')
    a, b = g, f - g.diff()*Poly(t, x)

    res, R = resultant(a, b, includePRS=True)
    res = Poly(res, t, composite=False)

    assert res, "BUG: resultant(%s, %s) cannot be zero" % (a, b)

    R_map, H = {}, []

    for r in R:
        R_map[r.degree()] = r

    def _include_sign(c, sqf):
        if c.is_extended_real and (c < 0) == True:
            h, k = sqf[0]
            c_poly = c.as_poly(h.gens)
            sqf[0] = h*c_poly, k

    C, res_sqf = res.sqf_list()
    _include_sign(C, res_sqf)

    for q, i in res_sqf:
        _, q = q.primitive()

        if g.degree() == i:
            H.append((g, q))
        else:
            h = R_map[i]
            h_lc = Poly(h.LC(), t, field=True)

            c, h_lc_sqf = h_lc.sqf_list(all=True)
            _include_sign(c, h_lc_sqf)

            for a, j in h_lc_sqf:
                h = h.quo(Poly(a.gcd(q)**j, x))

            inv, coeffs = h_lc.invert(q), [S.One]

            for coeff in h.coeffs()[1:]:
                coeff = coeff.as_poly(inv.gens)
                T = (inv*coeff).rem(q)
                coeffs.append(T.as_expr())

            h = Poly(dict(list(zip(h.monoms(), coeffs))), x)

            H.append((h, q))

    return H


def log_to_atan(f, g):
    """
    Convert complex logarithms to real arctangents.

    Explanation
    ===========

    Given a real field K and polynomials f and g in K[x], with g != 0,
    returns a sum h of arctangents of polynomials in K[x], such that:

                   dh   d         f + I g
                   -- = -- I log( ------- )
                   dx   dx        f - I g

    Examples
    ========

        >>> from sympy.integrals.rationaltools import log_to_atan
        >>> from sympy.abc import x
        >>> from sympy import Poly, sqrt, S
        >>> log_to_atan(Poly(x, x, domain='ZZ'), Poly(1, x, domain='ZZ'))
        2*atan(x)
        >>> log_to_atan(Poly(x + S(1)/2, x, domain='QQ'),
        ... Poly(sqrt(3)/2, x, domain='EX'))
        2*atan(2*sqrt(3)*x/3 + sqrt(3)/3)

    See Also
    ========

    log_to_real
    """
    if f.degree() < g.degree():
        f, g = -g, f

    f = f.to_field()
    g = g.to_field()

    p, q = f.div(g)

    if q.is_zero:
        return 2*atan(p.as_expr())
    else:
        s, t, h = g.gcdex(-f)
        u = (f*s + g*t).quo(h)
        A = 2*atan(u.as_expr())

        return A + log_to_atan(s, t)


def _get_real_roots(f, x):
    """get real roots of f if possible"""
    rs = roots(f, filter='R')

    try:
        num_roots = f.count_roots()
    except DomainError:
        return rs
    else:
        if len(rs) == num_roots:
            return rs
        else:
            return None


def log_to_real(h, q, x, t):
    r"""
    Convert complex logarithms to real functions.

    Explanation
    ===========

    Given real field K and polynomials h in K[t,x] and q in K[t],
    returns real function f such that:
                          ___
                  df   d  \  `
                  -- = --  )  a log(h(a, x))
                  dx   dx /__,
                         a | q(a) = 0

    Examples
    ========

        >>> from sympy.integrals.rationaltools import log_to_real
        >>> from sympy.abc import x, y
        >>> from sympy import Poly, S
        >>> log_to_real(Poly(x + 3*y/2 + S(1)/2, x, domain='QQ[y]'),
        ... Poly(3*y**2 + 1, y, domain='ZZ'), x, y)
        2*sqrt(3)*atan(2*sqrt(3)*x/3 + sqrt(3)/3)/3
        >>> log_to_real(Poly(x**2 - 1, x, domain='ZZ'),
        ... Poly(-2*y + 1, y, domain='ZZ'), x, y)
        log(x**2 - 1)/2

    See Also
    ========

    log_to_atan
    """
    from sympy.simplify.radsimp import collect
    u, v = symbols('u,v', cls=Dummy)

    H = h.as_expr().xreplace({t: u + I*v}).expand()
    Q = q.as_expr().xreplace({t: u + I*v}).expand()

    H_map = collect(H, I, evaluate=False)
    Q_map = collect(Q, I, evaluate=False)

    a, b = H_map.get(S.One, S.Zero), H_map.get(I, S.Zero)
    c, d = Q_map.get(S.One, S.Zero), Q_map.get(I, S.Zero)

    R = Poly(resultant(c, d, v), u)

    R_u = _get_real_roots(R, u)

    if R_u is None:
        return None

    result = S.Zero

    for r_u in R_u.keys():
        C = Poly(c.xreplace({u: r_u}), v)
        if not C:
            # t was split into real and imaginary parts
            # and denom Q(u, v) = c + I*d. We just found
            # that c(r_u) is 0 so the roots are in d
            C = Poly(d.xreplace({u: r_u}), v)
            # we were going to reject roots from C that
            # did not set d to zero, but since we are now
            # using C = d and c is already 0, there is
            # nothing to check
            d = S.Zero

        R_v = _get_real_roots(C, v)

        if R_v is None:
            return None

        R_v_paired = [] # take one from each pair of conjugate roots
        for r_v in R_v:
            if r_v not in R_v_paired and -r_v not in R_v_paired:
                if r_v.is_negative or r_v.could_extract_minus_sign():
                    R_v_paired.append(-r_v)
                elif not r_v.is_zero:
                    R_v_paired.append(r_v)

        for r_v in R_v_paired:

            D = d.xreplace({u: r_u, v: r_v})

            if D.evalf(chop=True) != 0:
                continue

            A = Poly(a.xreplace({u: r_u, v: r_v}), x)
            B = Poly(b.xreplace({u: r_u, v: r_v}), x)

            AB = (A**2 + B**2).as_expr()

            result += r_u*log(AB) + r_v*log_to_atan(A, B)

    R_q = _get_real_roots(q, t)

    if R_q is None:
        return None

    for r in R_q.keys():
        result += r*log(h.as_expr().subs(t, r))

    return result

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\print_page_options.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.


from typing import TYPE_CHECKING, List, Optional, Type

if TYPE_CHECKING:
    from typing import Literal, TypedDict

    Orientation = Literal["portrait", "landscape"]

    class _MarginOpts(TypedDict, total=False):
        left: float
        right: float
        top: float
        bottom: float

    class _PageOpts(TypedDict, total=False):
        width: float
        height: float

    class _PrintOpts(TypedDict, total=False):
        margin: _MarginOpts
        page: _PageOpts
        background: bool
        orientation: Orientation
        scale: float
        shrinkToFit: bool
        pageRanges: List[str]

else:
    from typing import Any, Dict

    Orientation = str
    _MarginOpts = _PageOpts = _PrintOpts = Dict[str, Any]


class _PageSettingsDescriptor:
    """Descriptor which validates `height` and 'width' of page."""

    def __init__(self, name):
        self.name = name

    def __get__(self, obj, cls) -> Optional[float]:
        return obj._page.get(self.name, None)

    def __set__(self, obj, value) -> None:
        getattr(obj, "_validate_num_property")(self.name, value)
        obj._page[self.name] = value
        obj._print_options["page"] = obj._page


class _MarginSettingsDescriptor:
    """Descriptor which validates below attributes.

    - top
    - bottom
    - left
    - right
    """

    def __init__(self, name):
        self.name = name

    def __get__(self, obj, cls) -> Optional[float]:
        return obj._margin.get(self.name, None)

    def __set__(self, obj, value) -> None:
        getattr(obj, "_validate_num_property")(f"Margin {self.name}", value)
        obj._margin[self.name] = value
        obj._print_options["margin"] = obj._margin


class _ScaleDescriptor:
    """Scale descriptor which validates scale."""

    def __init__(self, name):
        self.name = name

    def __get__(self, obj, cls) -> Optional[float]:
        return obj._print_options.get(self.name)

    def __set__(self, obj, value) -> None:
        getattr(obj, "_validate_num_property")(self.name, value)
        if value < 0.1 or value > 2:
            raise ValueError("Value of scale should be between 0.1 and 2")
        obj._print_options[self.name] = value


class _PageOrientationDescriptor:
    """PageOrientation descriptor which validates orientation of page."""

    ORIENTATION_VALUES = ["portrait", "landscape"]

    def __init__(self, name):
        self.name = name

    def __get__(self, obj, cls) -> Optional[Orientation]:
        return obj._print_options.get(self.name, None)

    def __set__(self, obj, value) -> None:
        if value not in self.ORIENTATION_VALUES:
            raise ValueError(f"Orientation value must be one of {self.ORIENTATION_VALUES}")
        obj._print_options[self.name] = value


class _ValidateTypeDescriptor:
    """Base Class Descriptor which validates type of any subclass attribute."""

    def __init__(self, name, expected_type: Type):
        self.name = name
        self.expected_type = expected_type

    def __get__(self, obj, cls):
        return obj._print_options.get(self.name, None)

    def __set__(self, obj, value) -> None:
        if not isinstance(value, self.expected_type):
            raise ValueError(f"{self.name} should be of type {self.expected_type.__name__}")
        obj._print_options[self.name] = value


class _ValidateBackGround(_ValidateTypeDescriptor):
    """Expected type of background attribute."""

    def __init__(self, name):
        super().__init__(name, bool)


class _ValidateShrinkToFit(_ValidateTypeDescriptor):
    """Expected type of shrink to fit attribute."""

    def __init__(self, name):
        super().__init__(name, bool)


class _ValidatePageRanges(_ValidateTypeDescriptor):
    """Expected type of page ranges attribute."""

    def __init__(self, name):
        super().__init__(name, list)


class PrintOptions:
    page_height = _PageSettingsDescriptor("height")
    """Gets and Sets page_height:

    Usage
    -----
    - Get
        - `self.page_height`
    - Set
        - `self.page_height` = `value`

    Parameters:
    -----------
    `value`: `float`

    Returns:
    --------
    - Get
        - `Optional[float]`
    - Set
        - `None`
    """

    page_width = _PageSettingsDescriptor("width")
    """Gets and Sets page_width:

    Usage:
    ------
    - Get
        - `self.page_width`
    - Set
        - `self.page_width` = `value`

    Parameters:
    -----------
    `value`: `float`

    Returns:
    --------
    - Get
        - `Optional[float]`
    - Set
        - `None`
    """

    margin_top = _MarginSettingsDescriptor("top")
    """Gets and Sets margin_top:

    Usage:
    ------
    - Get
        - `self.margin_top`
    - Set
        - `self.margin_top` = `value`

    Parameters:
    -----------
    `value`: `float`

    Returns:
    --------
    - Get
        - `Optional[float]`
    - Set
        - `None`
    """

    margin_bottom = _MarginSettingsDescriptor("bottom")
    """Gets and Sets margin_bottom:

    Usage:
    ------
    - Get
        - `self.margin_bottom`
    - Set
        - `self.margin_bottom` = `value`

    Parameters:
    -----------
    `value`: `float`

    Returns:
    --------
    - Get
        - `Optional[float]`
    - Set
        - `None`
    """

    margin_left = _MarginSettingsDescriptor("left")
    """Gets and Sets margin_left:

    Usage:
    ------
    - Get
        - `self.margin_left`
    - Set
        - `self.margin_left` = `value`

    Parameters:
    -----------
    `value`: `float`

    Returns:
    --------
    - Get
        - `Optional[float]`
    - Set
        - `None`
    """

    margin_right = _MarginSettingsDescriptor("right")
    """Gets and Sets margin_right:

    Usage:
    ------
    - Get
        - `self.margin_right`
    - Set
        - `self.margin_right` = `value`

    Parameters:
    -----------
    `value`: `float`

    Returns:
    --------
    - Get
        - `Optional[float]`
    - Set
        - `None`
    """

    scale = _ScaleDescriptor("scale")
    """Gets and Sets scale:

    Usage:
    ------
    - Get
        - `self.scale`
    - Set
        - `self.scale` = `value`

    Parameters:
    -----------
    `value`: `float`

    Returns:
    --------
    - Get
        - `Optional[float]`
    - Set
        - `None`
    """

    orientation = _PageOrientationDescriptor("orientation")
    """Gets and Sets orientation:

    Usage:
    ------
    - Get
        - `self.orientation`
    - Set
        - `self.orientation` = `value`

    Parameters:
    -----------
    `value`: `Orientation`

    Returns:
    --------
    - Get
        - `Optional[Orientation]`
    - Set
        - `None`
    """

    background = _ValidateBackGround("background")
    """Gets and Sets background:

    Usage:
    ------
    - Get
        - `self.background`
    - Set
        - `self.background` = `value`

    Parameters:
    -----------
    `value`: `bool`

    Returns:
    --------
    - Get
        - `Optional[bool]`
    - Set
        - `None`
    """

    shrink_to_fit = _ValidateShrinkToFit("shrinkToFit")
    """Gets and Sets shrink_to_fit:

    Usage:
    ------
    - Get
        - `self.shrink_to_fit`
    - Set
        - `self.shrink_to_fit` = `value`

    Parameters:
    -----------
    `value`: `bool`

    Returns:
    --------
    - Get
        - `Optional[bool]`
    - Set
        - `None`
    """

    page_ranges = _ValidatePageRanges("pageRanges")
    """Gets and Sets page_ranges:

    Usage:
    ------
    - Get
        - `self.page_ranges`
    - Set
        - `self.page_ranges` = `value`

    Parameters:
    -----------
    `value`: ` List[str]`

    Returns:
    --------
    - Get
        - `Optional[List[str]]`
    - Set
        - `None`
    """
    # Reference for predefined page size constants: https://www.agooddaytoprint.com/page/paper-size-chart-faq
    A4 = {"height": 29.7, "width": 21.0}  # size in cm
    LEGAL = {"height": 35.56, "width": 21.59}  # size in cm
    LETTER = {"height": 27.94, "width": 21.59}  # size in cm
    TABLOID = {"height": 43.18, "width": 27.94}  # size in cm

    def __init__(self) -> None:
        self._print_options: _PrintOpts = {}
        self._page: _PageOpts = {
            "height": PrintOptions.A4["height"],
            "width": PrintOptions.A4["width"],
        }  # Default page size set to A4
        self._margin: _MarginOpts = {}

    def to_dict(self) -> _PrintOpts:
        """:Returns: A hash of print options configured."""
        return self._print_options

    def set_page_size(self, page_size: dict) -> None:
        """Sets the page size to predefined or custom dimensions.

        Parameters:
        -----------
        page_size: dict
        A dictionary containing `height` and `width` as keys with respective values.

        Example:
        --------
        self.set_page_size(PageSize.A4)  # A4 predefined size
        self.set_page_size({"height": 15.0, "width": 20.0})  # Custom size in cm
        """
        self._validate_num_property("height", page_size["height"])
        self._validate_num_property("width", page_size["width"])
        self._page["height"] = page_size["height"]
        self._page["width"] = page_size["width"]
        self._print_options["page"] = self._page

    def _validate_num_property(self, property_name: str, value: float) -> None:
        """Helper function to validate some of the properties."""
        if not isinstance(value, (int, float)):
            raise ValueError(f"{property_name} should be an integer or a float")

        if value < 0:
            raise ValueError(f"{property_name} cannot be less than 0")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\layout.py ===
from abc import ABC, abstractmethod
from itertools import islice
from operator import itemgetter
from threading import RLock
from typing import (
    TYPE_CHECKING,
    Dict,
    Iterable,
    List,
    NamedTuple,
    Optional,
    Sequence,
    Tuple,
    Union,
)

from ._ratio import ratio_resolve
from .align import Align
from .console import Console, ConsoleOptions, RenderableType, RenderResult
from .highlighter import ReprHighlighter
from .panel import Panel
from .pretty import Pretty
from .region import Region
from .repr import Result, rich_repr
from .segment import Segment
from .style import StyleType

if TYPE_CHECKING:
    from pip._vendor.rich.tree import Tree


class LayoutRender(NamedTuple):
    """An individual layout render."""

    region: Region
    render: List[List[Segment]]


RegionMap = Dict["Layout", Region]
RenderMap = Dict["Layout", LayoutRender]


class LayoutError(Exception):
    """Layout related error."""


class NoSplitter(LayoutError):
    """Requested splitter does not exist."""


class _Placeholder:
    """An internal renderable used as a Layout placeholder."""

    highlighter = ReprHighlighter()

    def __init__(self, layout: "Layout", style: StyleType = "") -> None:
        self.layout = layout
        self.style = style

    def __rich_console__(
        self, console: Console, options: ConsoleOptions
    ) -> RenderResult:
        width = options.max_width
        height = options.height or options.size.height
        layout = self.layout
        title = (
            f"{layout.name!r} ({width} x {height})"
            if layout.name
            else f"({width} x {height})"
        )
        yield Panel(
            Align.center(Pretty(layout), vertical="middle"),
            style=self.style,
            title=self.highlighter(title),
            border_style="blue",
            height=height,
        )


class Splitter(ABC):
    """Base class for a splitter."""

    name: str = ""

    @abstractmethod
    def get_tree_icon(self) -> str:
        """Get the icon (emoji) used in layout.tree"""

    @abstractmethod
    def divide(
        self, children: Sequence["Layout"], region: Region
    ) -> Iterable[Tuple["Layout", Region]]:
        """Divide a region amongst several child layouts.

        Args:
            children (Sequence(Layout)): A number of child layouts.
            region (Region): A rectangular region to divide.
        """


class RowSplitter(Splitter):
    """Split a layout region in to rows."""

    name = "row"

    def get_tree_icon(self) -> str:
        return "[layout.tree.row]⬌"

    def divide(
        self, children: Sequence["Layout"], region: Region
    ) -> Iterable[Tuple["Layout", Region]]:
        x, y, width, height = region
        render_widths = ratio_resolve(width, children)
        offset = 0
        _Region = Region
        for child, child_width in zip(children, render_widths):
            yield child, _Region(x + offset, y, child_width, height)
            offset += child_width


class ColumnSplitter(Splitter):
    """Split a layout region in to columns."""

    name = "column"

    def get_tree_icon(self) -> str:
        return "[layout.tree.column]⬍"

    def divide(
        self, children: Sequence["Layout"], region: Region
    ) -> Iterable[Tuple["Layout", Region]]:
        x, y, width, height = region
        render_heights = ratio_resolve(height, children)
        offset = 0
        _Region = Region
        for child, child_height in zip(children, render_heights):
            yield child, _Region(x, y + offset, width, child_height)
            offset += child_height


@rich_repr
class Layout:
    """A renderable to divide a fixed height in to rows or columns.

    Args:
        renderable (RenderableType, optional): Renderable content, or None for placeholder. Defaults to None.
        name (str, optional): Optional identifier for Layout. Defaults to None.
        size (int, optional): Optional fixed size of layout. Defaults to None.
        minimum_size (int, optional): Minimum size of layout. Defaults to 1.
        ratio (int, optional): Optional ratio for flexible layout. Defaults to 1.
        visible (bool, optional): Visibility of layout. Defaults to True.
    """

    splitters = {"row": RowSplitter, "column": ColumnSplitter}

    def __init__(
        self,
        renderable: Optional[RenderableType] = None,
        *,
        name: Optional[str] = None,
        size: Optional[int] = None,
        minimum_size: int = 1,
        ratio: int = 1,
        visible: bool = True,
    ) -> None:
        self._renderable = renderable or _Placeholder(self)
        self.size = size
        self.minimum_size = minimum_size
        self.ratio = ratio
        self.name = name
        self.visible = visible
        self.splitter: Splitter = self.splitters["column"]()
        self._children: List[Layout] = []
        self._render_map: RenderMap = {}
        self._lock = RLock()

    def __rich_repr__(self) -> Result:
        yield "name", self.name, None
        yield "size", self.size, None
        yield "minimum_size", self.minimum_size, 1
        yield "ratio", self.ratio, 1

    @property
    def renderable(self) -> RenderableType:
        """Layout renderable."""
        return self if self._children else self._renderable

    @property
    def children(self) -> List["Layout"]:
        """Gets (visible) layout children."""
        return [child for child in self._children if child.visible]

    @property
    def map(self) -> RenderMap:
        """Get a map of the last render."""
        return self._render_map

    def get(self, name: str) -> Optional["Layout"]:
        """Get a named layout, or None if it doesn't exist.

        Args:
            name (str): Name of layout.

        Returns:
            Optional[Layout]: Layout instance or None if no layout was found.
        """
        if self.name == name:
            return self
        else:
            for child in self._children:
                named_layout = child.get(name)
                if named_layout is not None:
                    return named_layout
        return None

    def __getitem__(self, name: str) -> "Layout":
        layout = self.get(name)
        if layout is None:
            raise KeyError(f"No layout with name {name!r}")
        return layout

    @property
    def tree(self) -> "Tree":
        """Get a tree renderable to show layout structure."""
        from pip._vendor.rich.styled import Styled
        from pip._vendor.rich.table import Table
        from pip._vendor.rich.tree import Tree

        def summary(layout: "Layout") -> Table:
            icon = layout.splitter.get_tree_icon()

            table = Table.grid(padding=(0, 1, 0, 0))

            text: RenderableType = (
                Pretty(layout) if layout.visible else Styled(Pretty(layout), "dim")
            )
            table.add_row(icon, text)
            _summary = table
            return _summary

        layout = self
        tree = Tree(
            summary(layout),
            guide_style=f"layout.tree.{layout.splitter.name}",
            highlight=True,
        )

        def recurse(tree: "Tree", layout: "Layout") -> None:
            for child in layout._children:
                recurse(
                    tree.add(
                        summary(child),
                        guide_style=f"layout.tree.{child.splitter.name}",
                    ),
                    child,
                )

        recurse(tree, self)
        return tree

    def split(
        self,
        *layouts: Union["Layout", RenderableType],
        splitter: Union[Splitter, str] = "column",
    ) -> None:
        """Split the layout in to multiple sub-layouts.

        Args:
            *layouts (Layout): Positional arguments should be (sub) Layout instances.
            splitter (Union[Splitter, str]): Splitter instance or name of splitter.
        """
        _layouts = [
            layout if isinstance(layout, Layout) else Layout(layout)
            for layout in layouts
        ]
        try:
            self.splitter = (
                splitter
                if isinstance(splitter, Splitter)
                else self.splitters[splitter]()
            )
        except KeyError:
            raise NoSplitter(f"No splitter called {splitter!r}")
        self._children[:] = _layouts

    def add_split(self, *layouts: Union["Layout", RenderableType]) -> None:
        """Add a new layout(s) to existing split.

        Args:
            *layouts (Union[Layout, RenderableType]): Positional arguments should be renderables or (sub) Layout instances.

        """
        _layouts = (
            layout if isinstance(layout, Layout) else Layout(layout)
            for layout in layouts
        )
        self._children.extend(_layouts)

    def split_row(self, *layouts: Union["Layout", RenderableType]) -> None:
        """Split the layout in to a row (layouts side by side).

        Args:
            *layouts (Layout): Positional arguments should be (sub) Layout instances.
        """
        self.split(*layouts, splitter="row")

    def split_column(self, *layouts: Union["Layout", RenderableType]) -> None:
        """Split the layout in to a column (layouts stacked on top of each other).

        Args:
            *layouts (Layout): Positional arguments should be (sub) Layout instances.
        """
        self.split(*layouts, splitter="column")

    def unsplit(self) -> None:
        """Reset splits to initial state."""
        del self._children[:]

    def update(self, renderable: RenderableType) -> None:
        """Update renderable.

        Args:
            renderable (RenderableType): New renderable object.
        """
        with self._lock:
            self._renderable = renderable

    def refresh_screen(self, console: "Console", layout_name: str) -> None:
        """Refresh a sub-layout.

        Args:
            console (Console): Console instance where Layout is to be rendered.
            layout_name (str): Name of layout.
        """
        with self._lock:
            layout = self[layout_name]
            region, _lines = self._render_map[layout]
            (x, y, width, height) = region
            lines = console.render_lines(
                layout, console.options.update_dimensions(width, height)
            )
            self._render_map[layout] = LayoutRender(region, lines)
            console.update_screen_lines(lines, x, y)

    def _make_region_map(self, width: int, height: int) -> RegionMap:
        """Create a dict that maps layout on to Region."""
        stack: List[Tuple[Layout, Region]] = [(self, Region(0, 0, width, height))]
        push = stack.append
        pop = stack.pop
        layout_regions: List[Tuple[Layout, Region]] = []
        append_layout_region = layout_regions.append
        while stack:
            append_layout_region(pop())
            layout, region = layout_regions[-1]
            children = layout.children
            if children:
                for child_and_region in layout.splitter.divide(children, region):
                    push(child_and_region)

        region_map = {
            layout: region
            for layout, region in sorted(layout_regions, key=itemgetter(1))
        }
        return region_map

    def render(self, console: Console, options: ConsoleOptions) -> RenderMap:
        """Render the sub_layouts.

        Args:
            console (Console): Console instance.
            options (ConsoleOptions): Console options.

        Returns:
            RenderMap: A dict that maps Layout on to a tuple of Region, lines
        """
        render_width = options.max_width
        render_height = options.height or console.height
        region_map = self._make_region_map(render_width, render_height)
        layout_regions = [
            (layout, region)
            for layout, region in region_map.items()
            if not layout.children
        ]
        render_map: Dict["Layout", "LayoutRender"] = {}
        render_lines = console.render_lines
        update_dimensions = options.update_dimensions

        for layout, region in layout_regions:
            lines = render_lines(
                layout.renderable, update_dimensions(region.width, region.height)
            )
            render_map[layout] = LayoutRender(region, lines)
        return render_map

    def __rich_console__(
        self, console: Console, options: ConsoleOptions
    ) -> RenderResult:
        with self._lock:
            width = options.max_width or console.width
            height = options.height or console.height
            render_map = self.render(console, options.update_dimensions(width, height))
            self._render_map = render_map
            layout_lines: List[List[Segment]] = [[] for _ in range(height)]
            _islice = islice
            for region, lines in render_map.values():
                _x, y, _layout_width, layout_height = region
                for row, line in zip(
                    _islice(layout_lines, y, y + layout_height), lines
                ):
                    row.extend(line)

            new_line = Segment.line()
            for layout_row in layout_lines:
                yield from layout_row
                yield new_line


if __name__ == "__main__":
    from pip._vendor.rich.console import Console

    console = Console()
    layout = Layout()

    layout.split_column(
        Layout(name="header", size=3),
        Layout(ratio=1, name="main"),
        Layout(size=10, name="footer"),
    )

    layout["main"].split_row(Layout(name="side"), Layout(name="body", ratio=2))

    layout["body"].split_row(Layout(name="content", ratio=2), Layout(name="s2"))

    layout["s2"].split_column(
        Layout(name="top"), Layout(name="middle"), Layout(name="bottom")
    )

    layout["side"].split_column(Layout(layout.tree, name="left1"), Layout(name="left2"))

    layout["content"].update("foo")

    console.print(layout)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\assumptions\handlers\order.py ===
"""
Handlers related to order relations: positive, negative, etc.
"""

from sympy.assumptions import Q, ask
from sympy.core import Add, Basic, Expr, Mul, Pow, S
from sympy.core.logic import fuzzy_not, fuzzy_and, fuzzy_or
from sympy.core.numbers import E, ImaginaryUnit, NaN, I, pi
from sympy.functions import Abs, acos, acot, asin, atan, exp, factorial, log
from sympy.matrices import Determinant, Trace
from sympy.matrices.expressions.matexpr import MatrixElement

from sympy.multipledispatch import MDNotImplementedError

from ..predicates.order import (NegativePredicate, NonNegativePredicate,
    NonZeroPredicate, ZeroPredicate, NonPositivePredicate, PositivePredicate,
    ExtendedNegativePredicate, ExtendedNonNegativePredicate,
    ExtendedNonPositivePredicate, ExtendedNonZeroPredicate,
    ExtendedPositivePredicate,)


# NegativePredicate

def _NegativePredicate_number(expr, assumptions):
    r, i = expr.as_real_imag()

    if r == S.NaN or i == S.NaN:
        return None

    # If the imaginary part can symbolically be shown to be zero then
    # we just evaluate the real part; otherwise we evaluate the imaginary
    # part to see if it actually evaluates to zero and if it does then
    # we make the comparison between the real part and zero.
    if not i:
        r = r.evalf(2)
        if r._prec != 1:
            return r < 0
    else:
        i = i.evalf(2)
        if i._prec != 1:
            if i != 0:
                return False
            r = r.evalf(2)
            if r._prec != 1:
                return r < 0

@NegativePredicate.register(Basic)
def _(expr, assumptions):
    if expr.is_number:
        return _NegativePredicate_number(expr, assumptions)

@NegativePredicate.register(Expr)
def _(expr, assumptions):
    ret = expr.is_negative
    if ret is None:
        raise MDNotImplementedError
    return ret

@NegativePredicate.register(Add)
def _(expr, assumptions):
    """
    Positive + Positive -> Positive,
    Negative + Negative -> Negative
    """
    if expr.is_number:
        return _NegativePredicate_number(expr, assumptions)

    r = ask(Q.real(expr), assumptions)
    if r is not True:
        return r

    nonpos = 0
    for arg in expr.args:
        if ask(Q.negative(arg), assumptions) is not True:
            if ask(Q.positive(arg), assumptions) is False:
                nonpos += 1
            else:
                break
    else:
        if nonpos < len(expr.args):
            return True

@NegativePredicate.register(Mul)
def _(expr, assumptions):
    if expr.is_number:
        return _NegativePredicate_number(expr, assumptions)
    result = None
    for arg in expr.args:
        if result is None:
            result = False
        if ask(Q.negative(arg), assumptions):
            result = not result
        elif ask(Q.positive(arg), assumptions):
            pass
        else:
            return
    return result

@NegativePredicate.register(Pow)
def _(expr, assumptions):
    """
    Real ** Even -> NonNegative
    Real ** Odd  -> same_as_base
    NonNegative ** Positive -> NonNegative
    """
    if expr.base == E:
        # Exponential is always positive:
        if ask(Q.real(expr.exp), assumptions):
            return False
        return

    if expr.is_number:
        return _NegativePredicate_number(expr, assumptions)
    if ask(Q.real(expr.base), assumptions):
        if ask(Q.positive(expr.base), assumptions):
            if ask(Q.real(expr.exp), assumptions):
                return False
        if ask(Q.even(expr.exp), assumptions):
            return False
        if ask(Q.odd(expr.exp), assumptions):
            return ask(Q.negative(expr.base), assumptions)

@NegativePredicate.register_many(Abs, ImaginaryUnit)
def _(expr, assumptions):
    return False

@NegativePredicate.register(exp)
def _(expr, assumptions):
    if ask(Q.real(expr.exp), assumptions):
        return False
    raise MDNotImplementedError


# NonNegativePredicate

@NonNegativePredicate.register(Basic)
def _(expr, assumptions):
    if expr.is_number:
        notnegative = fuzzy_not(_NegativePredicate_number(expr, assumptions))
        if notnegative:
            return ask(Q.real(expr), assumptions)
        else:
            return notnegative

@NonNegativePredicate.register(Expr)
def _(expr, assumptions):
    ret = expr.is_nonnegative
    if ret is None:
        raise MDNotImplementedError
    return ret


# NonZeroPredicate

@NonZeroPredicate.register(Expr)
def _(expr, assumptions):
    ret = expr.is_nonzero
    if ret is None:
        raise MDNotImplementedError
    return ret

@NonZeroPredicate.register(Basic)
def _(expr, assumptions):
    if ask(Q.real(expr)) is False:
        return False
    if expr.is_number:
        # if there are no symbols just evalf
        i = expr.evalf(2)
        def nonz(i):
            if i._prec != 1:
                return i != 0
        return fuzzy_or(nonz(i) for i in i.as_real_imag())

@NonZeroPredicate.register(Add)
def _(expr, assumptions):
    if all(ask(Q.positive(x), assumptions) for x in expr.args) \
            or all(ask(Q.negative(x), assumptions) for x in expr.args):
        return True

@NonZeroPredicate.register(Mul)
def _(expr, assumptions):
    for arg in expr.args:
        result = ask(Q.nonzero(arg), assumptions)
        if result:
            continue
        return result
    return True

@NonZeroPredicate.register(Pow)
def _(expr, assumptions):
    return ask(Q.nonzero(expr.base), assumptions)

@NonZeroPredicate.register(Abs)
def _(expr, assumptions):
    return ask(Q.nonzero(expr.args[0]), assumptions)

@NonZeroPredicate.register(NaN)
def _(expr, assumptions):
    return None


# ZeroPredicate

@ZeroPredicate.register(Expr)
def _(expr, assumptions):
    ret = expr.is_zero
    if ret is None:
        raise MDNotImplementedError
    return ret

@ZeroPredicate.register(Basic)
def _(expr, assumptions):
    return fuzzy_and([fuzzy_not(ask(Q.nonzero(expr), assumptions)),
        ask(Q.real(expr), assumptions)])

@ZeroPredicate.register(Mul)
def _(expr, assumptions):
    # TODO: This should be deducible from the nonzero handler
    return fuzzy_or(ask(Q.zero(arg), assumptions) for arg in expr.args)


# NonPositivePredicate

@NonPositivePredicate.register(Expr)
def _(expr, assumptions):
    ret = expr.is_nonpositive
    if ret is None:
        raise MDNotImplementedError
    return ret

@NonPositivePredicate.register(Basic)
def _(expr, assumptions):
    if expr.is_number:
        notpositive = fuzzy_not(_PositivePredicate_number(expr, assumptions))
        if notpositive:
            return ask(Q.real(expr), assumptions)
        else:
            return notpositive


# PositivePredicate

def _PositivePredicate_number(expr, assumptions):
    r, i = expr.as_real_imag()
    # If the imaginary part can symbolically be shown to be zero then
    # we just evaluate the real part; otherwise we evaluate the imaginary
    # part to see if it actually evaluates to zero and if it does then
    # we make the comparison between the real part and zero.
    if not i:
        r = r.evalf(2)
        if r._prec != 1:
            return r > 0
    else:
        i = i.evalf(2)
        if i._prec != 1:
            if i != 0:
                return False
            r = r.evalf(2)
            if r._prec != 1:
                return r > 0

@PositivePredicate.register(Expr)
def _(expr, assumptions):
    ret = expr.is_positive
    if ret is None:
        raise MDNotImplementedError
    return ret

@PositivePredicate.register(Basic)
def _(expr, assumptions):
    if expr.is_number:
        return _PositivePredicate_number(expr, assumptions)

@PositivePredicate.register(Mul)
def _(expr, assumptions):
    if expr.is_number:
        return _PositivePredicate_number(expr, assumptions)
    result = True
    for arg in expr.args:
        if ask(Q.positive(arg), assumptions):
            continue
        elif ask(Q.negative(arg), assumptions):
            result = result ^ True
        else:
            return
    return result

@PositivePredicate.register(Add)
def _(expr, assumptions):
    if expr.is_number:
        return _PositivePredicate_number(expr, assumptions)

    r = ask(Q.real(expr), assumptions)
    if r is not True:
        return r

    nonneg = 0
    for arg in expr.args:
        if ask(Q.positive(arg), assumptions) is not True:
            if ask(Q.negative(arg), assumptions) is False:
                nonneg += 1
            else:
                break
    else:
        if nonneg < len(expr.args):
            return True

@PositivePredicate.register(Pow)
def _(expr, assumptions):
    if expr.base == E:
        if ask(Q.real(expr.exp), assumptions):
            return True
        if ask(Q.imaginary(expr.exp), assumptions):
            return ask(Q.even(expr.exp/(I*pi)), assumptions)
        return

    if expr.is_number:
        return _PositivePredicate_number(expr, assumptions)
    if ask(Q.positive(expr.base), assumptions):
        if ask(Q.real(expr.exp), assumptions):
            return True
    if ask(Q.negative(expr.base), assumptions):
        if ask(Q.even(expr.exp), assumptions):
            return True
        if ask(Q.odd(expr.exp), assumptions):
            return False

@PositivePredicate.register(exp)
def _(expr, assumptions):
    if ask(Q.real(expr.exp), assumptions):
        return True
    if ask(Q.imaginary(expr.exp), assumptions):
        return ask(Q.even(expr.exp/(I*pi)), assumptions)

@PositivePredicate.register(log)
def _(expr, assumptions):
    r = ask(Q.real(expr.args[0]), assumptions)
    if r is not True:
        return r
    if ask(Q.positive(expr.args[0] - 1), assumptions):
        return True
    if ask(Q.negative(expr.args[0] - 1), assumptions):
        return False

@PositivePredicate.register(factorial)
def _(expr, assumptions):
    x = expr.args[0]
    if ask(Q.integer(x) & Q.positive(x), assumptions):
            return True

@PositivePredicate.register(ImaginaryUnit)
def _(expr, assumptions):
    return False

@PositivePredicate.register(Abs)
def _(expr, assumptions):
    return ask(Q.nonzero(expr), assumptions)

@PositivePredicate.register(Trace)
def _(expr, assumptions):
    if ask(Q.positive_definite(expr.arg), assumptions):
        return True

@PositivePredicate.register(Determinant)
def _(expr, assumptions):
    if ask(Q.positive_definite(expr.arg), assumptions):
        return True

@PositivePredicate.register(MatrixElement)
def _(expr, assumptions):
    if (expr.i == expr.j
            and ask(Q.positive_definite(expr.parent), assumptions)):
        return True

@PositivePredicate.register(atan)
def _(expr, assumptions):
    return ask(Q.positive(expr.args[0]), assumptions)

@PositivePredicate.register(asin)
def _(expr, assumptions):
    x = expr.args[0]
    if ask(Q.positive(x) & Q.nonpositive(x - 1), assumptions):
        return True
    if ask(Q.negative(x) & Q.nonnegative(x + 1), assumptions):
        return False

@PositivePredicate.register(acos)
def _(expr, assumptions):
    x = expr.args[0]
    if ask(Q.nonpositive(x - 1) & Q.nonnegative(x + 1), assumptions):
        return True

@PositivePredicate.register(acot)
def _(expr, assumptions):
    return ask(Q.real(expr.args[0]), assumptions)

@PositivePredicate.register(NaN)
def _(expr, assumptions):
    return None


# ExtendedNegativePredicate

@ExtendedNegativePredicate.register(object)
def _(expr, assumptions):
    return ask(Q.negative(expr) | Q.negative_infinite(expr), assumptions)


# ExtendedPositivePredicate

@ExtendedPositivePredicate.register(object)
def _(expr, assumptions):
    return ask(Q.positive(expr) | Q.positive_infinite(expr), assumptions)


# ExtendedNonZeroPredicate

@ExtendedNonZeroPredicate.register(object)
def _(expr, assumptions):
    return ask(
        Q.negative_infinite(expr) | Q.negative(expr) | Q.positive(expr) | Q.positive_infinite(expr),
        assumptions)


# ExtendedNonPositivePredicate

@ExtendedNonPositivePredicate.register(object)
def _(expr, assumptions):
    return ask(
        Q.negative_infinite(expr) | Q.negative(expr) | Q.zero(expr),
        assumptions)


# ExtendedNonNegativePredicate

@ExtendedNonNegativePredicate.register(object)
def _(expr, assumptions):
    return ask(
        Q.zero(expr) | Q.positive(expr) | Q.positive_infinite(expr),
        assumptions)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\locations\__init__.py ===
import functools
import logging
import os
import pathlib
import sys
import sysconfig
from typing import Any, Dict, Optional

from pip._internal.models.scheme import SCHEME_KEYS, Scheme
from pip._internal.utils.compat import WINDOWS
from pip._internal.utils.deprecation import deprecated
from pip._internal.utils.virtualenv import running_under_virtualenv

from . import _sysconfig
from .base import (
    USER_CACHE_DIR,
    get_major_minor_version,
    get_src_prefix,
    is_osx_framework,
    site_packages,
    user_site,
)

__all__ = [
    "USER_CACHE_DIR",
    "get_bin_prefix",
    "get_bin_user",
    "get_major_minor_version",
    "get_platlib",
    "get_purelib",
    "get_scheme",
    "get_src_prefix",
    "site_packages",
    "user_site",
]


logger = logging.getLogger(__name__)


_PLATLIBDIR: str = getattr(sys, "platlibdir", "lib")

_USE_SYSCONFIG_DEFAULT = sys.version_info >= (3, 10)


def _should_use_sysconfig() -> bool:
    """This function determines the value of _USE_SYSCONFIG.

    By default, pip uses sysconfig on Python 3.10+.
    But Python distributors can override this decision by setting:
        sysconfig._PIP_USE_SYSCONFIG = True / False
    Rationale in https://github.com/pypa/pip/issues/10647

    This is a function for testability, but should be constant during any one
    run.
    """
    return bool(getattr(sysconfig, "_PIP_USE_SYSCONFIG", _USE_SYSCONFIG_DEFAULT))


_USE_SYSCONFIG = _should_use_sysconfig()

if not _USE_SYSCONFIG:
    # Import distutils lazily to avoid deprecation warnings,
    # but import it soon enough that it is in memory and available during
    # a pip reinstall.
    from . import _distutils

# Be noisy about incompatibilities if this platforms "should" be using
# sysconfig, but is explicitly opting out and using distutils instead.
if _USE_SYSCONFIG_DEFAULT and not _USE_SYSCONFIG:
    _MISMATCH_LEVEL = logging.WARNING
else:
    _MISMATCH_LEVEL = logging.DEBUG


def _looks_like_bpo_44860() -> bool:
    """The resolution to bpo-44860 will change this incorrect platlib.

    See <https://bugs.python.org/issue44860>.
    """
    from distutils.command.install import INSTALL_SCHEMES

    try:
        unix_user_platlib = INSTALL_SCHEMES["unix_user"]["platlib"]
    except KeyError:
        return False
    return unix_user_platlib == "$usersite"


def _looks_like_red_hat_patched_platlib_purelib(scheme: Dict[str, str]) -> bool:
    platlib = scheme["platlib"]
    if "/$platlibdir/" in platlib:
        platlib = platlib.replace("/$platlibdir/", f"/{_PLATLIBDIR}/")
    if "/lib64/" not in platlib:
        return False
    unpatched = platlib.replace("/lib64/", "/lib/")
    return unpatched.replace("$platbase/", "$base/") == scheme["purelib"]


@functools.lru_cache(maxsize=None)
def _looks_like_red_hat_lib() -> bool:
    """Red Hat patches platlib in unix_prefix and unix_home, but not purelib.

    This is the only way I can see to tell a Red Hat-patched Python.
    """
    from distutils.command.install import INSTALL_SCHEMES

    return all(
        k in INSTALL_SCHEMES
        and _looks_like_red_hat_patched_platlib_purelib(INSTALL_SCHEMES[k])
        for k in ("unix_prefix", "unix_home")
    )


@functools.lru_cache(maxsize=None)
def _looks_like_debian_scheme() -> bool:
    """Debian adds two additional schemes."""
    from distutils.command.install import INSTALL_SCHEMES

    return "deb_system" in INSTALL_SCHEMES and "unix_local" in INSTALL_SCHEMES


@functools.lru_cache(maxsize=None)
def _looks_like_red_hat_scheme() -> bool:
    """Red Hat patches ``sys.prefix`` and ``sys.exec_prefix``.

    Red Hat's ``00251-change-user-install-location.patch`` changes the install
    command's ``prefix`` and ``exec_prefix`` to append ``"/local"``. This is
    (fortunately?) done quite unconditionally, so we create a default command
    object without any configuration to detect this.
    """
    from distutils.command.install import install
    from distutils.dist import Distribution

    cmd: Any = install(Distribution())
    cmd.finalize_options()
    return (
        cmd.exec_prefix == f"{os.path.normpath(sys.exec_prefix)}/local"
        and cmd.prefix == f"{os.path.normpath(sys.prefix)}/local"
    )


@functools.lru_cache(maxsize=None)
def _looks_like_slackware_scheme() -> bool:
    """Slackware patches sysconfig but fails to patch distutils and site.

    Slackware changes sysconfig's user scheme to use ``"lib64"`` for the lib
    path, but does not do the same to the site module.
    """
    if user_site is None:  # User-site not available.
        return False
    try:
        paths = sysconfig.get_paths(scheme="posix_user", expand=False)
    except KeyError:  # User-site not available.
        return False
    return "/lib64/" in paths["purelib"] and "/lib64/" not in user_site


@functools.lru_cache(maxsize=None)
def _looks_like_msys2_mingw_scheme() -> bool:
    """MSYS2 patches distutils and sysconfig to use a UNIX-like scheme.

    However, MSYS2 incorrectly patches sysconfig ``nt`` scheme. The fix is
    likely going to be included in their 3.10 release, so we ignore the warning.
    See msys2/MINGW-packages#9319.

    MSYS2 MINGW's patch uses lowercase ``"lib"`` instead of the usual uppercase,
    and is missing the final ``"site-packages"``.
    """
    paths = sysconfig.get_paths("nt", expand=False)
    return all(
        "Lib" not in p and "lib" in p and not p.endswith("site-packages")
        for p in (paths[key] for key in ("platlib", "purelib"))
    )


@functools.lru_cache(maxsize=None)
def _warn_mismatched(old: pathlib.Path, new: pathlib.Path, *, key: str) -> None:
    issue_url = "https://github.com/pypa/pip/issues/10151"
    message = (
        "Value for %s does not match. Please report this to <%s>"
        "\ndistutils: %s"
        "\nsysconfig: %s"
    )
    logger.log(_MISMATCH_LEVEL, message, key, issue_url, old, new)


def _warn_if_mismatch(old: pathlib.Path, new: pathlib.Path, *, key: str) -> bool:
    if old == new:
        return False
    _warn_mismatched(old, new, key=key)
    return True


@functools.lru_cache(maxsize=None)
def _log_context(
    *,
    user: bool = False,
    home: Optional[str] = None,
    root: Optional[str] = None,
    prefix: Optional[str] = None,
) -> None:
    parts = [
        "Additional context:",
        "user = %r",
        "home = %r",
        "root = %r",
        "prefix = %r",
    ]

    logger.log(_MISMATCH_LEVEL, "\n".join(parts), user, home, root, prefix)


def get_scheme(
    dist_name: str,
    user: bool = False,
    home: Optional[str] = None,
    root: Optional[str] = None,
    isolated: bool = False,
    prefix: Optional[str] = None,
) -> Scheme:
    new = _sysconfig.get_scheme(
        dist_name,
        user=user,
        home=home,
        root=root,
        isolated=isolated,
        prefix=prefix,
    )
    if _USE_SYSCONFIG:
        return new

    old = _distutils.get_scheme(
        dist_name,
        user=user,
        home=home,
        root=root,
        isolated=isolated,
        prefix=prefix,
    )

    warning_contexts = []
    for k in SCHEME_KEYS:
        old_v = pathlib.Path(getattr(old, k))
        new_v = pathlib.Path(getattr(new, k))

        if old_v == new_v:
            continue

        # distutils incorrectly put PyPy packages under ``site-packages/python``
        # in the ``posix_home`` scheme, but PyPy devs said they expect the
        # directory name to be ``pypy`` instead. So we treat this as a bug fix
        # and not warn about it. See bpo-43307 and python/cpython#24628.
        skip_pypy_special_case = (
            sys.implementation.name == "pypy"
            and home is not None
            and k in ("platlib", "purelib")
            and old_v.parent == new_v.parent
            and old_v.name.startswith("python")
            and new_v.name.startswith("pypy")
        )
        if skip_pypy_special_case:
            continue

        # sysconfig's ``osx_framework_user`` does not include ``pythonX.Y`` in
        # the ``include`` value, but distutils's ``headers`` does. We'll let
        # CPython decide whether this is a bug or feature. See bpo-43948.
        skip_osx_framework_user_special_case = (
            user
            and is_osx_framework()
            and k == "headers"
            and old_v.parent.parent == new_v.parent
            and old_v.parent.name.startswith("python")
        )
        if skip_osx_framework_user_special_case:
            continue

        # On Red Hat and derived Linux distributions, distutils is patched to
        # use "lib64" instead of "lib" for platlib.
        if k == "platlib" and _looks_like_red_hat_lib():
            continue

        # On Python 3.9+, sysconfig's posix_user scheme sets platlib against
        # sys.platlibdir, but distutils's unix_user incorrectly coninutes
        # using the same $usersite for both platlib and purelib. This creates a
        # mismatch when sys.platlibdir is not "lib".
        skip_bpo_44860 = (
            user
            and k == "platlib"
            and not WINDOWS
            and _PLATLIBDIR != "lib"
            and _looks_like_bpo_44860()
        )
        if skip_bpo_44860:
            continue

        # Slackware incorrectly patches posix_user to use lib64 instead of lib,
        # but not usersite to match the location.
        skip_slackware_user_scheme = (
            user
            and k in ("platlib", "purelib")
            and not WINDOWS
            and _looks_like_slackware_scheme()
        )
        if skip_slackware_user_scheme:
            continue

        # Both Debian and Red Hat patch Python to place the system site under
        # /usr/local instead of /usr. Debian also places lib in dist-packages
        # instead of site-packages, but the /usr/local check should cover it.
        skip_linux_system_special_case = (
            not (user or home or prefix or running_under_virtualenv())
            and old_v.parts[1:3] == ("usr", "local")
            and len(new_v.parts) > 1
            and new_v.parts[1] == "usr"
            and (len(new_v.parts) < 3 or new_v.parts[2] != "local")
            and (_looks_like_red_hat_scheme() or _looks_like_debian_scheme())
        )
        if skip_linux_system_special_case:
            continue

        # MSYS2 MINGW's sysconfig patch does not include the "site-packages"
        # part of the path. This is incorrect and will be fixed in MSYS.
        skip_msys2_mingw_bug = (
            WINDOWS and k in ("platlib", "purelib") and _looks_like_msys2_mingw_scheme()
        )
        if skip_msys2_mingw_bug:
            continue

        # CPython's POSIX install script invokes pip (via ensurepip) against the
        # interpreter located in the source tree, not the install site. This
        # triggers special logic in sysconfig that's not present in distutils.
        # https://github.com/python/cpython/blob/8c21941ddaf/Lib/sysconfig.py#L178-L194
        skip_cpython_build = (
            sysconfig.is_python_build(check_home=True)
            and not WINDOWS
            and k in ("headers", "include", "platinclude")
        )
        if skip_cpython_build:
            continue

        warning_contexts.append((old_v, new_v, f"scheme.{k}"))

    if not warning_contexts:
        return old

    # Check if this path mismatch is caused by distutils config files. Those
    # files will no longer work once we switch to sysconfig, so this raises a
    # deprecation message for them.
    default_old = _distutils.distutils_scheme(
        dist_name,
        user,
        home,
        root,
        isolated,
        prefix,
        ignore_config_files=True,
    )
    if any(default_old[k] != getattr(old, k) for k in SCHEME_KEYS):
        deprecated(
            reason=(
                "Configuring installation scheme with distutils config files "
                "is deprecated and will no longer work in the near future. If you "
                "are using a Homebrew or Linuxbrew Python, please see discussion "
                "at https://github.com/Homebrew/homebrew-core/issues/76621"
            ),
            replacement=None,
            gone_in=None,
        )
        return old

    # Post warnings about this mismatch so user can report them back.
    for old_v, new_v, key in warning_contexts:
        _warn_mismatched(old_v, new_v, key=key)
    _log_context(user=user, home=home, root=root, prefix=prefix)

    return old


def get_bin_prefix() -> str:
    new = _sysconfig.get_bin_prefix()
    if _USE_SYSCONFIG:
        return new

    old = _distutils.get_bin_prefix()
    if _warn_if_mismatch(pathlib.Path(old), pathlib.Path(new), key="bin_prefix"):
        _log_context()
    return old


def get_bin_user() -> str:
    return _sysconfig.get_scheme("", user=True).scripts


def _looks_like_deb_system_dist_packages(value: str) -> bool:
    """Check if the value is Debian's APT-controlled dist-packages.

    Debian's ``distutils.sysconfig.get_python_lib()`` implementation returns the
    default package path controlled by APT, but does not patch ``sysconfig`` to
    do the same. This is similar to the bug worked around in ``get_scheme()``,
    but here the default is ``deb_system`` instead of ``unix_local``. Ultimately
    we can't do anything about this Debian bug, and this detection allows us to
    skip the warning when needed.
    """
    if not _looks_like_debian_scheme():
        return False
    if value == "/usr/lib/python3/dist-packages":
        return True
    return False


def get_purelib() -> str:
    """Return the default pure-Python lib location."""
    new = _sysconfig.get_purelib()
    if _USE_SYSCONFIG:
        return new

    old = _distutils.get_purelib()
    if _looks_like_deb_system_dist_packages(old):
        return old
    if _warn_if_mismatch(pathlib.Path(old), pathlib.Path(new), key="purelib"):
        _log_context()
    return old


def get_platlib() -> str:
    """Return the default platform-shared lib location."""
    new = _sysconfig.get_platlib()
    if _USE_SYSCONFIG:
        return new

    from . import _distutils

    old = _distutils.get_platlib()
    if _looks_like_deb_system_dist_packages(old):
        return old
    if _warn_if_mismatch(pathlib.Path(old), pathlib.Path(new), key="platlib"):
        _log_context()
    return old

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\werkzeug\middleware\lint.py ===
"""
WSGI Protocol Linter
====================

This module provides a middleware that performs sanity checks on the
behavior of the WSGI server and application. It checks that the
:pep:`3333` WSGI spec is properly implemented. It also warns on some
common HTTP errors such as non-empty responses for 304 status codes.

.. autoclass:: LintMiddleware

:copyright: 2007 Pallets
:license: BSD-3-Clause
"""

from __future__ import annotations

import typing as t
from types import TracebackType
from urllib.parse import urlparse
from warnings import warn

from ..datastructures import Headers
from ..http import is_entity_header
from ..wsgi import FileWrapper

if t.TYPE_CHECKING:
    from _typeshed.wsgi import StartResponse
    from _typeshed.wsgi import WSGIApplication
    from _typeshed.wsgi import WSGIEnvironment


class WSGIWarning(Warning):
    """Warning class for WSGI warnings."""


class HTTPWarning(Warning):
    """Warning class for HTTP warnings."""


def check_type(context: str, obj: object, need: type = str) -> None:
    if type(obj) is not need:
        warn(
            f"{context!r} requires {need.__name__!r}, got {type(obj).__name__!r}.",
            WSGIWarning,
            stacklevel=3,
        )


class InputStream:
    def __init__(self, stream: t.IO[bytes]) -> None:
        self._stream = stream

    def read(self, *args: t.Any) -> bytes:
        if len(args) == 0:
            warn(
                "WSGI does not guarantee an EOF marker on the input stream, thus making"
                " calls to 'wsgi.input.read()' unsafe. Conforming servers may never"
                " return from this call.",
                WSGIWarning,
                stacklevel=2,
            )
        elif len(args) != 1:
            warn(
                "Too many parameters passed to 'wsgi.input.read()'.",
                WSGIWarning,
                stacklevel=2,
            )
        return self._stream.read(*args)

    def readline(self, *args: t.Any) -> bytes:
        if len(args) == 0:
            warn(
                "Calls to 'wsgi.input.readline()' without arguments are unsafe. Use"
                " 'wsgi.input.read()' instead.",
                WSGIWarning,
                stacklevel=2,
            )
        elif len(args) == 1:
            warn(
                "'wsgi.input.readline()' was called with a size hint. WSGI does not"
                " support this, although it's available on all major servers.",
                WSGIWarning,
                stacklevel=2,
            )
        else:
            raise TypeError("Too many arguments passed to 'wsgi.input.readline()'.")
        return self._stream.readline(*args)

    def __iter__(self) -> t.Iterator[bytes]:
        try:
            return iter(self._stream)
        except TypeError:
            warn("'wsgi.input' is not iterable.", WSGIWarning, stacklevel=2)
            return iter(())

    def close(self) -> None:
        warn("The application closed the input stream!", WSGIWarning, stacklevel=2)
        self._stream.close()


class ErrorStream:
    def __init__(self, stream: t.IO[str]) -> None:
        self._stream = stream

    def write(self, s: str) -> None:
        check_type("wsgi.error.write()", s, str)
        self._stream.write(s)

    def flush(self) -> None:
        self._stream.flush()

    def writelines(self, seq: t.Iterable[str]) -> None:
        for line in seq:
            self.write(line)

    def close(self) -> None:
        warn("The application closed the error stream!", WSGIWarning, stacklevel=2)
        self._stream.close()


class GuardedWrite:
    def __init__(self, write: t.Callable[[bytes], object], chunks: list[int]) -> None:
        self._write = write
        self._chunks = chunks

    def __call__(self, s: bytes) -> None:
        check_type("write()", s, bytes)
        self._write(s)
        self._chunks.append(len(s))


class GuardedIterator:
    def __init__(
        self,
        iterator: t.Iterable[bytes],
        headers_set: tuple[int, Headers],
        chunks: list[int],
    ) -> None:
        self._iterator = iterator
        self._next = iter(iterator).__next__
        self.closed = False
        self.headers_set = headers_set
        self.chunks = chunks

    def __iter__(self) -> GuardedIterator:
        return self

    def __next__(self) -> bytes:
        if self.closed:
            warn("Iterated over closed 'app_iter'.", WSGIWarning, stacklevel=2)

        rv = self._next()

        if not self.headers_set:
            warn(
                "The application returned before it started the response.",
                WSGIWarning,
                stacklevel=2,
            )

        check_type("application iterator items", rv, bytes)
        self.chunks.append(len(rv))
        return rv

    def close(self) -> None:
        self.closed = True

        if hasattr(self._iterator, "close"):
            self._iterator.close()

        if self.headers_set:
            status_code, headers = self.headers_set
            bytes_sent = sum(self.chunks)
            content_length = headers.get("content-length", type=int)

            if status_code == 304:
                for key, _value in headers:
                    key = key.lower()
                    if key not in ("expires", "content-location") and is_entity_header(
                        key
                    ):
                        warn(
                            f"Entity header {key!r} found in 304 response.",
                            HTTPWarning,
                            stacklevel=2,
                        )
                if bytes_sent:
                    warn(
                        "304 responses must not have a body.",
                        HTTPWarning,
                        stacklevel=2,
                    )
            elif 100 <= status_code < 200 or status_code == 204:
                if content_length != 0:
                    warn(
                        f"{status_code} responses must have an empty content length.",
                        HTTPWarning,
                        stacklevel=2,
                    )
                if bytes_sent:
                    warn(
                        f"{status_code} responses must not have a body.",
                        HTTPWarning,
                        stacklevel=2,
                    )
            elif content_length is not None and content_length != bytes_sent:
                warn(
                    "Content-Length and the number of bytes sent to the"
                    " client do not match.",
                    WSGIWarning,
                    stacklevel=2,
                )

    def __del__(self) -> None:
        if not self.closed:
            try:
                warn(
                    "Iterator was garbage collected before it was closed.",
                    WSGIWarning,
                    stacklevel=2,
                )
            except Exception:
                pass


class LintMiddleware:
    """Warns about common errors in the WSGI and HTTP behavior of the
    server and wrapped application. Some of the issues it checks are:

    -   invalid status codes
    -   non-bytes sent to the WSGI server
    -   strings returned from the WSGI application
    -   non-empty conditional responses
    -   unquoted etags
    -   relative URLs in the Location header
    -   unsafe calls to wsgi.input
    -   unclosed iterators

    Error information is emitted using the :mod:`warnings` module.

    :param app: The WSGI application to wrap.

    .. code-block:: python

        from werkzeug.middleware.lint import LintMiddleware
        app = LintMiddleware(app)
    """

    def __init__(self, app: WSGIApplication) -> None:
        self.app = app

    def check_environ(self, environ: WSGIEnvironment) -> None:
        if type(environ) is not dict:  # noqa: E721
            warn(
                "WSGI environment is not a standard Python dict.",
                WSGIWarning,
                stacklevel=4,
            )
        for key in (
            "REQUEST_METHOD",
            "SERVER_NAME",
            "SERVER_PORT",
            "wsgi.version",
            "wsgi.input",
            "wsgi.errors",
            "wsgi.multithread",
            "wsgi.multiprocess",
            "wsgi.run_once",
        ):
            if key not in environ:
                warn(
                    f"Required environment key {key!r} not found",
                    WSGIWarning,
                    stacklevel=3,
                )
        if environ["wsgi.version"] != (1, 0):
            warn("Environ is not a WSGI 1.0 environ.", WSGIWarning, stacklevel=3)

        script_name = environ.get("SCRIPT_NAME", "")
        path_info = environ.get("PATH_INFO", "")

        if script_name and script_name[0] != "/":
            warn(
                f"'SCRIPT_NAME' does not start with a slash: {script_name!r}",
                WSGIWarning,
                stacklevel=3,
            )

        if path_info and path_info[0] != "/":
            warn(
                f"'PATH_INFO' does not start with a slash: {path_info!r}",
                WSGIWarning,
                stacklevel=3,
            )

    def check_start_response(
        self,
        status: str,
        headers: list[tuple[str, str]],
        exc_info: None | (tuple[type[BaseException], BaseException, TracebackType]),
    ) -> tuple[int, Headers]:
        check_type("status", status, str)
        status_code_str = status.split(None, 1)[0]

        if len(status_code_str) != 3 or not status_code_str.isdecimal():
            warn("Status code must be three digits.", WSGIWarning, stacklevel=3)

        if len(status) < 4 or status[3] != " ":
            warn(
                f"Invalid value for status {status!r}. Valid status strings are three"
                " digits, a space and a status explanation.",
                WSGIWarning,
                stacklevel=3,
            )

        status_code = int(status_code_str)

        if status_code < 100:
            warn("Status code < 100 detected.", WSGIWarning, stacklevel=3)

        if type(headers) is not list:  # noqa: E721
            warn("Header list is not a list.", WSGIWarning, stacklevel=3)

        for item in headers:
            if type(item) is not tuple or len(item) != 2:
                warn("Header items must be 2-item tuples.", WSGIWarning, stacklevel=3)
            name, value = item
            if type(name) is not str or type(value) is not str:  # noqa: E721
                warn(
                    "Header keys and values must be strings.", WSGIWarning, stacklevel=3
                )
            if name.lower() == "status":
                warn(
                    "The status header is not supported due to"
                    " conflicts with the CGI spec.",
                    WSGIWarning,
                    stacklevel=3,
                )

        if exc_info is not None and not isinstance(exc_info, tuple):
            warn("Invalid value for exc_info.", WSGIWarning, stacklevel=3)

        headers_obj = Headers(headers)
        self.check_headers(headers_obj)

        return status_code, headers_obj

    def check_headers(self, headers: Headers) -> None:
        etag = headers.get("etag")

        if etag is not None:
            if etag.startswith(("W/", "w/")):
                if etag.startswith("w/"):
                    warn(
                        "Weak etag indicator should be upper case.",
                        HTTPWarning,
                        stacklevel=4,
                    )

                etag = etag[2:]

            if not (etag[:1] == etag[-1:] == '"'):
                warn("Unquoted etag emitted.", HTTPWarning, stacklevel=4)

        location = headers.get("location")

        if location is not None:
            if not urlparse(location).netloc:
                warn(
                    "Absolute URLs required for location header.",
                    HTTPWarning,
                    stacklevel=4,
                )

    def check_iterator(self, app_iter: t.Iterable[bytes]) -> None:
        if isinstance(app_iter, str):
            warn(
                "The application returned a string. The response will send one"
                " character at a time to the client, which will kill performance."
                " Return a list or iterable instead.",
                WSGIWarning,
                stacklevel=3,
            )

    def __call__(self, *args: t.Any, **kwargs: t.Any) -> t.Iterable[bytes]:
        if len(args) != 2:
            warn("A WSGI app takes two arguments.", WSGIWarning, stacklevel=2)

        if kwargs:
            warn(
                "A WSGI app does not take keyword arguments.", WSGIWarning, stacklevel=2
            )

        environ: WSGIEnvironment = args[0]
        start_response: StartResponse = args[1]

        self.check_environ(environ)
        environ["wsgi.input"] = InputStream(environ["wsgi.input"])
        environ["wsgi.errors"] = ErrorStream(environ["wsgi.errors"])

        # Hook our own file wrapper in so that applications will always
        # iterate to the end and we can check the content length.
        environ["wsgi.file_wrapper"] = FileWrapper

        headers_set: list[t.Any] = []
        chunks: list[int] = []

        def checking_start_response(
            *args: t.Any, **kwargs: t.Any
        ) -> t.Callable[[bytes], None]:
            if len(args) not in {2, 3}:
                warn(
                    f"Invalid number of arguments: {len(args)}, expected 2 or 3.",
                    WSGIWarning,
                    stacklevel=2,
                )

            if kwargs:
                warn(
                    "'start_response' does not take keyword arguments.",
                    WSGIWarning,
                    stacklevel=2,
                )

            status: str = args[0]
            headers: list[tuple[str, str]] = args[1]
            exc_info: (
                None | (tuple[type[BaseException], BaseException, TracebackType])
            ) = args[2] if len(args) == 3 else None

            headers_set[:] = self.check_start_response(status, headers, exc_info)
            return GuardedWrite(start_response(status, headers, exc_info), chunks)

        app_iter = self.app(environ, t.cast("StartResponse", checking_start_response))
        self.check_iterator(app_iter)
        return GuardedIterator(
            app_iter, t.cast(tuple[int, Headers], headers_set), chunks
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\workbook\workbook.py ===
# Copyright (c) 2010-2024 openpyxl

"""Workbook is the top-level container for all document information."""
from copy import copy

from openpyxl.compat import deprecated
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet.copier import WorksheetCopy

from openpyxl.utils import quote_sheetname
from openpyxl.utils.indexed_list import IndexedList
from openpyxl.utils.datetime  import WINDOWS_EPOCH, MAC_EPOCH
from openpyxl.utils.exceptions import ReadOnlyWorkbookException

from openpyxl.writer.excel import save_workbook

from openpyxl.styles.cell_style import StyleArray
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.styles.differential import DifferentialStyleList
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import DEFAULT_BORDER
from openpyxl.styles.fills import DEFAULT_EMPTY_FILL, DEFAULT_GRAY_FILL
from openpyxl.styles.fonts import DEFAULT_FONT
from openpyxl.styles.protection import Protection
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.styles.named_styles import NamedStyleList
from openpyxl.styles.table import TableStyleList

from openpyxl.chartsheet import Chartsheet
from .defined_name import DefinedName, DefinedNameDict
from openpyxl.packaging.core import DocumentProperties
from openpyxl.packaging.custom import CustomPropertyList
from openpyxl.packaging.relationship import RelationshipList
from .child import _WorkbookChild
from .protection import DocumentSecurity
from .properties import CalcProperties
from .views import BookView


from openpyxl.xml.constants import (
    XLSM,
    XLSX,
    XLTM,
    XLTX
)

INTEGER_TYPES = (int,)

class Workbook:
    """Workbook is the container for all other parts of the document."""

    _read_only = False
    _data_only = False
    template = False
    path = "/xl/workbook.xml"

    def __init__(self,
                 write_only=False,
                 iso_dates=False,
                 ):
        self._sheets = []
        self._pivots = []
        self._active_sheet_index = 0
        self.defined_names = DefinedNameDict()
        self._external_links = []
        self.properties = DocumentProperties()
        self.custom_doc_props = CustomPropertyList()
        self.security = DocumentSecurity()
        self.__write_only = write_only
        self.shared_strings = IndexedList()

        self._setup_styles()

        self.loaded_theme = None
        self.vba_archive = None
        self.is_template = False
        self.code_name = None
        self.epoch = WINDOWS_EPOCH
        self.encoding = "utf-8"
        self.iso_dates = iso_dates

        if not self.write_only:
            self._sheets.append(Worksheet(self))

        self.rels = RelationshipList()
        self.calculation = CalcProperties()
        self.views = [BookView()]


    def _setup_styles(self):
        """Bootstrap styles"""

        self._fonts = IndexedList()
        self._fonts.add(DEFAULT_FONT)

        self._alignments = IndexedList([Alignment()])

        self._borders = IndexedList()
        self._borders.add(DEFAULT_BORDER)

        self._fills = IndexedList()
        self._fills.add(DEFAULT_EMPTY_FILL)
        self._fills.add(DEFAULT_GRAY_FILL)

        self._number_formats = IndexedList()
        self._date_formats = {}
        self._timedelta_formats = {}

        self._protections = IndexedList([Protection()])

        self._colors = COLOR_INDEX
        self._cell_styles = IndexedList([StyleArray()])
        self._named_styles = NamedStyleList()
        self.add_named_style(NamedStyle(font=copy(DEFAULT_FONT), border=copy(DEFAULT_BORDER), builtinId=0))
        self._table_styles = TableStyleList()
        self._differential_styles = DifferentialStyleList()


    @property
    def epoch(self):
        if self._epoch == WINDOWS_EPOCH:
            return WINDOWS_EPOCH
        return MAC_EPOCH


    @epoch.setter
    def epoch(self, value):
        if value not in (WINDOWS_EPOCH, MAC_EPOCH):
            raise ValueError("The epoch must be either 1900 or 1904")
        self._epoch = value


    @property
    def read_only(self):
        return self._read_only

    @property
    def data_only(self):
        return self._data_only

    @property
    def write_only(self):
        return self.__write_only


    @property
    def excel_base_date(self):
        return self.epoch

    @property
    def active(self):
        """Get the currently active sheet or None

        :type: :class:`openpyxl.worksheet.worksheet.Worksheet`
        """
        try:
            return self._sheets[self._active_sheet_index]
        except IndexError:
            pass

    @active.setter
    def active(self, value):
        """Set the active sheet"""
        if not isinstance(value, (_WorkbookChild, INTEGER_TYPES)):
            raise TypeError("Value must be either a worksheet, chartsheet or numerical index")
        if isinstance(value, INTEGER_TYPES):
            self._active_sheet_index = value
            return
            #if self._sheets and 0 <= value < len(self._sheets):
                #value = self._sheets[value]
            #else:
                #raise ValueError("Sheet index is outside the range of possible values", value)
        if value not in self._sheets:
            raise ValueError("Worksheet is not in the workbook")
        if value.sheet_state != "visible":
            raise ValueError("Only visible sheets can be made active")

        idx = self._sheets.index(value)
        self._active_sheet_index = idx


    def create_sheet(self, title=None, index=None):
        """Create a worksheet (at an optional index).

        :param title: optional title of the sheet
        :type title: str
        :param index: optional position at which the sheet will be inserted
        :type index: int

        """
        if self.read_only:
            raise ReadOnlyWorkbookException('Cannot create new sheet in a read-only workbook')

        if self.write_only :
            new_ws = WriteOnlyWorksheet(parent=self, title=title)
        else:
            new_ws = Worksheet(parent=self, title=title)

        self._add_sheet(sheet=new_ws, index=index)
        return new_ws


    def _add_sheet(self, sheet, index=None):
        """Add an worksheet (at an optional index)."""

        if not isinstance(sheet, (Worksheet, WriteOnlyWorksheet, Chartsheet)):
            raise TypeError("Cannot be added to a workbook")

        if sheet.parent != self:
            raise ValueError("You cannot add worksheets from another workbook.")

        if index is None:
            self._sheets.append(sheet)
        else:
            self._sheets.insert(index, sheet)


    def move_sheet(self, sheet, offset=0):
        """
        Move a sheet or sheetname
        """
        if not isinstance(sheet, Worksheet):
            sheet = self[sheet]
        idx = self._sheets.index(sheet)
        del self._sheets[idx]
        new_pos = idx + offset
        self._sheets.insert(new_pos, sheet)


    def remove(self, worksheet):
        """Remove `worksheet` from this workbook."""
        idx = self._sheets.index(worksheet)
        self._sheets.remove(worksheet)


    @deprecated("Use wb.remove(worksheet) or del wb[sheetname]")
    def remove_sheet(self, worksheet):
        """Remove `worksheet` from this workbook."""
        self.remove(worksheet)


    def create_chartsheet(self, title=None, index=None):
        if self.read_only:
            raise ReadOnlyWorkbookException("Cannot create new sheet in a read-only workbook")
        cs = Chartsheet(parent=self, title=title)

        self._add_sheet(cs, index)
        return cs


    @deprecated("Use wb[sheetname]")
    def get_sheet_by_name(self, name):
        """Returns a worksheet by its name.

        :param name: the name of the worksheet to look for
        :type name: string

        """
        return self[name]

    def __contains__(self, key):
        return key in self.sheetnames


    def index(self, worksheet):
        """Return the index of a worksheet."""
        return self.worksheets.index(worksheet)


    @deprecated("Use wb.index(worksheet)")
    def get_index(self, worksheet):
        """Return the index of the worksheet."""
        return self.index(worksheet)

    def __getitem__(self, key):
        """Returns a worksheet by its name.

        :param name: the name of the worksheet to look for
        :type name: string

        """
        for sheet in self.worksheets + self.chartsheets:
            if sheet.title == key:
                return sheet
        raise KeyError("Worksheet {0} does not exist.".format(key))

    def __delitem__(self, key):
        sheet = self[key]
        self.remove(sheet)

    def __iter__(self):
        return iter(self.worksheets)


    @deprecated("Use wb.sheetnames")
    def get_sheet_names(self):
        return self.sheetnames

    @property
    def worksheets(self):
        """A list of sheets in this workbook

        :type: list of :class:`openpyxl.worksheet.worksheet.Worksheet`
        """
        return [s for s in self._sheets if isinstance(s, (Worksheet, ReadOnlyWorksheet, WriteOnlyWorksheet))]

    @property
    def chartsheets(self):
        """A list of Chartsheets in this workbook

        :type: list of :class:`openpyxl.chartsheet.chartsheet.Chartsheet`
        """
        return [s for s in self._sheets if isinstance(s, Chartsheet)]

    @property
    def sheetnames(self):
        """Returns the list of the names of worksheets in this workbook.

        Names are returned in the worksheets order.

        :type: list of strings

        """
        return [s.title for s in self._sheets]


    @deprecated("Assign scoped named ranges directly to worksheets or global ones to the workbook. Deprecated in 3.1")
    def create_named_range(self, name, worksheet=None, value=None, scope=None):
        """Create a new named_range on a worksheet

        """
        defn = DefinedName(name=name)
        if worksheet is not None:
            defn.value = "{0}!{1}".format(quote_sheetname(worksheet.title), value)
        else:
            defn.value = value

        self.defined_names[name] = defn


    def add_named_style(self, style):
        """
        Add a named style
        """
        self._named_styles.append(style)
        style.bind(self)


    @property
    def named_styles(self):
        """
        List available named styles
        """
        return self._named_styles.names


    @property
    def mime_type(self):
        """
        The mime type is determined by whether a workbook is a template or
        not and whether it contains macros or not. Excel requires the file
        extension to match but openpyxl does not enforce this.

        """
        ct = self.template and XLTX or XLSX
        if self.vba_archive:
            ct = self.template and XLTM or XLSM
        return ct


    def save(self, filename):
        """Save the current workbook under the given `filename`.
        Use this function instead of using an `ExcelWriter`.

        .. warning::
            When creating your workbook using `write_only` set to True,
            you will only be able to call this function once. Subsequent attempts to
            modify or save the file will raise an :class:`openpyxl.shared.exc.WorkbookAlreadySaved` exception.
        """
        if self.read_only:
            raise TypeError("""Workbook is read-only""")
        if self.write_only and not self.worksheets:
            self.create_sheet()
        save_workbook(self, filename)


    @property
    def style_names(self):
        """
        List of named styles
        """
        return [s.name for s in self._named_styles]


    def copy_worksheet(self, from_worksheet):
        """Copy an existing worksheet in the current workbook

        .. warning::
            This function cannot copy worksheets between workbooks.
            worksheets can only be copied within the workbook that they belong

        :param from_worksheet: the worksheet to be copied from
        :return: copy of the initial worksheet
        """
        if self.__write_only or self._read_only:
            raise ValueError("Cannot copy worksheets in read-only or write-only mode")

        new_title = u"{0} Copy".format(from_worksheet.title)
        to_worksheet = self.create_sheet(title=new_title)
        cp = WorksheetCopy(source_worksheet=from_worksheet, target_worksheet=to_worksheet)
        cp.copy_worksheet()
        return to_worksheet


    def close(self):
        """
        Close workbook file if open. Only affects read-only and write-only modes.
        """
        if hasattr(self, '_archive'):
            self._archive.close()


    def _duplicate_name(self, name):
        """
        Check for duplicate name in defined name list and table list of each worksheet.
        Names are not case sensitive.
        """
        name = name.lower()
        for sheet in self.worksheets:
            for t in sheet.tables:
                if name == t.lower():
                    return True

        if name in self.defined_names:
            return True


# === atelier-kyo-manager/.venv_backup\Lib\site-packages\idna\core.py ===
import bisect
import re
import unicodedata
from typing import Optional, Union

from . import idnadata
from .intranges import intranges_contain

_virama_combining_class = 9
_alabel_prefix = b"xn--"
_unicode_dots_re = re.compile("[\u002e\u3002\uff0e\uff61]")


class IDNAError(UnicodeError):
    """Base exception for all IDNA-encoding related problems"""

    pass


class IDNABidiError(IDNAError):
    """Exception when bidirectional requirements are not satisfied"""

    pass


class InvalidCodepoint(IDNAError):
    """Exception when a disallowed or unallocated codepoint is used"""

    pass


class InvalidCodepointContext(IDNAError):
    """Exception when the codepoint is not valid in the context it is used"""

    pass


def _combining_class(cp: int) -> int:
    v = unicodedata.combining(chr(cp))
    if v == 0:
        if not unicodedata.name(chr(cp)):
            raise ValueError("Unknown character in unicodedata")
    return v


def _is_script(cp: str, script: str) -> bool:
    return intranges_contain(ord(cp), idnadata.scripts[script])


def _punycode(s: str) -> bytes:
    return s.encode("punycode")


def _unot(s: int) -> str:
    return "U+{:04X}".format(s)


def valid_label_length(label: Union[bytes, str]) -> bool:
    if len(label) > 63:
        return False
    return True


def valid_string_length(label: Union[bytes, str], trailing_dot: bool) -> bool:
    if len(label) > (254 if trailing_dot else 253):
        return False
    return True


def check_bidi(label: str, check_ltr: bool = False) -> bool:
    # Bidi rules should only be applied if string contains RTL characters
    bidi_label = False
    for idx, cp in enumerate(label, 1):
        direction = unicodedata.bidirectional(cp)
        if direction == "":
            # String likely comes from a newer version of Unicode
            raise IDNABidiError("Unknown directionality in label {} at position {}".format(repr(label), idx))
        if direction in ["R", "AL", "AN"]:
            bidi_label = True
    if not bidi_label and not check_ltr:
        return True

    # Bidi rule 1
    direction = unicodedata.bidirectional(label[0])
    if direction in ["R", "AL"]:
        rtl = True
    elif direction == "L":
        rtl = False
    else:
        raise IDNABidiError("First codepoint in label {} must be directionality L, R or AL".format(repr(label)))

    valid_ending = False
    number_type: Optional[str] = None
    for idx, cp in enumerate(label, 1):
        direction = unicodedata.bidirectional(cp)

        if rtl:
            # Bidi rule 2
            if direction not in [
                "R",
                "AL",
                "AN",
                "EN",
                "ES",
                "CS",
                "ET",
                "ON",
                "BN",
                "NSM",
            ]:
                raise IDNABidiError("Invalid direction for codepoint at position {} in a right-to-left label".format(idx))
            # Bidi rule 3
            if direction in ["R", "AL", "EN", "AN"]:
                valid_ending = True
            elif direction != "NSM":
                valid_ending = False
            # Bidi rule 4
            if direction in ["AN", "EN"]:
                if not number_type:
                    number_type = direction
                else:
                    if number_type != direction:
                        raise IDNABidiError("Can not mix numeral types in a right-to-left label")
        else:
            # Bidi rule 5
            if direction not in ["L", "EN", "ES", "CS", "ET", "ON", "BN", "NSM"]:
                raise IDNABidiError("Invalid direction for codepoint at position {} in a left-to-right label".format(idx))
            # Bidi rule 6
            if direction in ["L", "EN"]:
                valid_ending = True
            elif direction != "NSM":
                valid_ending = False

    if not valid_ending:
        raise IDNABidiError("Label ends with illegal codepoint directionality")

    return True


def check_initial_combiner(label: str) -> bool:
    if unicodedata.category(label[0])[0] == "M":
        raise IDNAError("Label begins with an illegal combining character")
    return True


def check_hyphen_ok(label: str) -> bool:
    if label[2:4] == "--":
        raise IDNAError("Label has disallowed hyphens in 3rd and 4th position")
    if label[0] == "-" or label[-1] == "-":
        raise IDNAError("Label must not start or end with a hyphen")
    return True


def check_nfc(label: str) -> None:
    if unicodedata.normalize("NFC", label) != label:
        raise IDNAError("Label must be in Normalization Form C")


def valid_contextj(label: str, pos: int) -> bool:
    cp_value = ord(label[pos])

    if cp_value == 0x200C:
        if pos > 0:
            if _combining_class(ord(label[pos - 1])) == _virama_combining_class:
                return True

        ok = False
        for i in range(pos - 1, -1, -1):
            joining_type = idnadata.joining_types.get(ord(label[i]))
            if joining_type == ord("T"):
                continue
            elif joining_type in [ord("L"), ord("D")]:
                ok = True
                break
            else:
                break

        if not ok:
            return False

        ok = False
        for i in range(pos + 1, len(label)):
            joining_type = idnadata.joining_types.get(ord(label[i]))
            if joining_type == ord("T"):
                continue
            elif joining_type in [ord("R"), ord("D")]:
                ok = True
                break
            else:
                break
        return ok

    if cp_value == 0x200D:
        if pos > 0:
            if _combining_class(ord(label[pos - 1])) == _virama_combining_class:
                return True
        return False

    else:
        return False


def valid_contexto(label: str, pos: int, exception: bool = False) -> bool:
    cp_value = ord(label[pos])

    if cp_value == 0x00B7:
        if 0 < pos < len(label) - 1:
            if ord(label[pos - 1]) == 0x006C and ord(label[pos + 1]) == 0x006C:
                return True
        return False

    elif cp_value == 0x0375:
        if pos < len(label) - 1 and len(label) > 1:
            return _is_script(label[pos + 1], "Greek")
        return False

    elif cp_value == 0x05F3 or cp_value == 0x05F4:
        if pos > 0:
            return _is_script(label[pos - 1], "Hebrew")
        return False

    elif cp_value == 0x30FB:
        for cp in label:
            if cp == "\u30fb":
                continue
            if _is_script(cp, "Hiragana") or _is_script(cp, "Katakana") or _is_script(cp, "Han"):
                return True
        return False

    elif 0x660 <= cp_value <= 0x669:
        for cp in label:
            if 0x6F0 <= ord(cp) <= 0x06F9:
                return False
        return True

    elif 0x6F0 <= cp_value <= 0x6F9:
        for cp in label:
            if 0x660 <= ord(cp) <= 0x0669:
                return False
        return True

    return False


def check_label(label: Union[str, bytes, bytearray]) -> None:
    if isinstance(label, (bytes, bytearray)):
        label = label.decode("utf-8")
    if len(label) == 0:
        raise IDNAError("Empty Label")

    check_nfc(label)
    check_hyphen_ok(label)
    check_initial_combiner(label)

    for pos, cp in enumerate(label):
        cp_value = ord(cp)
        if intranges_contain(cp_value, idnadata.codepoint_classes["PVALID"]):
            continue
        elif intranges_contain(cp_value, idnadata.codepoint_classes["CONTEXTJ"]):
            try:
                if not valid_contextj(label, pos):
                    raise InvalidCodepointContext(
                        "Joiner {} not allowed at position {} in {}".format(_unot(cp_value), pos + 1, repr(label))
                    )
            except ValueError:
                raise IDNAError(
                    "Unknown codepoint adjacent to joiner {} at position {} in {}".format(
                        _unot(cp_value), pos + 1, repr(label)
                    )
                )
        elif intranges_contain(cp_value, idnadata.codepoint_classes["CONTEXTO"]):
            if not valid_contexto(label, pos):
                raise InvalidCodepointContext(
                    "Codepoint {} not allowed at position {} in {}".format(_unot(cp_value), pos + 1, repr(label))
                )
        else:
            raise InvalidCodepoint(
                "Codepoint {} at position {} of {} not allowed".format(_unot(cp_value), pos + 1, repr(label))
            )

    check_bidi(label)


def alabel(label: str) -> bytes:
    try:
        label_bytes = label.encode("ascii")
        ulabel(label_bytes)
        if not valid_label_length(label_bytes):
            raise IDNAError("Label too long")
        return label_bytes
    except UnicodeEncodeError:
        pass

    check_label(label)
    label_bytes = _alabel_prefix + _punycode(label)

    if not valid_label_length(label_bytes):
        raise IDNAError("Label too long")

    return label_bytes


def ulabel(label: Union[str, bytes, bytearray]) -> str:
    if not isinstance(label, (bytes, bytearray)):
        try:
            label_bytes = label.encode("ascii")
        except UnicodeEncodeError:
            check_label(label)
            return label
    else:
        label_bytes = label

    label_bytes = label_bytes.lower()
    if label_bytes.startswith(_alabel_prefix):
        label_bytes = label_bytes[len(_alabel_prefix) :]
        if not label_bytes:
            raise IDNAError("Malformed A-label, no Punycode eligible content found")
        if label_bytes.decode("ascii")[-1] == "-":
            raise IDNAError("A-label must not end with a hyphen")
    else:
        check_label(label_bytes)
        return label_bytes.decode("ascii")

    try:
        label = label_bytes.decode("punycode")
    except UnicodeError:
        raise IDNAError("Invalid A-label")
    check_label(label)
    return label


def uts46_remap(domain: str, std3_rules: bool = True, transitional: bool = False) -> str:
    """Re-map the characters in the string according to UTS46 processing."""
    from .uts46data import uts46data

    output = ""

    for pos, char in enumerate(domain):
        code_point = ord(char)
        try:
            uts46row = uts46data[code_point if code_point < 256 else bisect.bisect_left(uts46data, (code_point, "Z")) - 1]
            status = uts46row[1]
            replacement: Optional[str] = None
            if len(uts46row) == 3:
                replacement = uts46row[2]
            if (
                status == "V"
                or (status == "D" and not transitional)
                or (status == "3" and not std3_rules and replacement is None)
            ):
                output += char
            elif replacement is not None and (
                status == "M" or (status == "3" and not std3_rules) or (status == "D" and transitional)
            ):
                output += replacement
            elif status != "I":
                raise IndexError()
        except IndexError:
            raise InvalidCodepoint(
                "Codepoint {} not allowed at position {} in {}".format(_unot(code_point), pos + 1, repr(domain))
            )

    return unicodedata.normalize("NFC", output)


def encode(
    s: Union[str, bytes, bytearray],
    strict: bool = False,
    uts46: bool = False,
    std3_rules: bool = False,
    transitional: bool = False,
) -> bytes:
    if not isinstance(s, str):
        try:
            s = str(s, "ascii")
        except UnicodeDecodeError:
            raise IDNAError("should pass a unicode string to the function rather than a byte string.")
    if uts46:
        s = uts46_remap(s, std3_rules, transitional)
    trailing_dot = False
    result = []
    if strict:
        labels = s.split(".")
    else:
        labels = _unicode_dots_re.split(s)
    if not labels or labels == [""]:
        raise IDNAError("Empty domain")
    if labels[-1] == "":
        del labels[-1]
        trailing_dot = True
    for label in labels:
        s = alabel(label)
        if s:
            result.append(s)
        else:
            raise IDNAError("Empty label")
    if trailing_dot:
        result.append(b"")
    s = b".".join(result)
    if not valid_string_length(s, trailing_dot):
        raise IDNAError("Domain too long")
    return s


def decode(
    s: Union[str, bytes, bytearray],
    strict: bool = False,
    uts46: bool = False,
    std3_rules: bool = False,
) -> str:
    try:
        if not isinstance(s, str):
            s = str(s, "ascii")
    except UnicodeDecodeError:
        raise IDNAError("Invalid ASCII in A-label")
    if uts46:
        s = uts46_remap(s, std3_rules, False)
    trailing_dot = False
    result = []
    if not strict:
        labels = _unicode_dots_re.split(s)
    else:
        labels = s.split(".")
    if not labels or labels == [""]:
        raise IDNAError("Empty domain")
    if not labels[-1]:
        del labels[-1]
        trailing_dot = True
    for label in labels:
        s = ulabel(label)
        if s:
            result.append(s)
        else:
            raise IDNAError("Empty label")
    if trailing_dot:
        result.append("")
    return ".".join(result)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\t5\t5_decoder.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.  See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------

import logging
import os
import tempfile
from pathlib import Path

import numpy
import onnx
import torch
from io_binding_helper import TypeHelper
from onnx_model import OnnxModel
from past_helper import PastKeyValuesHelper
from t5_encoder import T5EncoderInputs
from torch_onnx_export_helper import torch_onnx_export
from transformers import MT5Config, T5Config

from onnxruntime import InferenceSession

logger = logging.getLogger(__name__)


class T5DecoderInit(torch.nn.Module):
    """A T5 decoder with LM head to create initial past key values.
    This model is only called once during starting decoding.
    """

    def __init__(
        self,
        decoder: torch.nn.Module,
        lm_head: torch.nn.Module,
        config: T5Config | MT5Config,
        decoder_start_token_id: int | None = None,
    ):
        super().__init__()
        self.decoder = decoder
        self.lm_head = lm_head
        self.config = config
        self.decoder_start_token_id = (
            decoder_start_token_id if decoder_start_token_id is not None else self.config.decoder_start_token_id
        )
        self.tie_word_embeddings = (
            self.config.tie_word_embeddings if hasattr(self.config, "tie_word_embeddings") else True
        )

    def forward(
        self,
        decoder_input_ids: torch.Tensor,
        encoder_attention_mask: torch.Tensor,
        encoder_hidden_states: torch.FloatTensor,
    ):
        if decoder_input_ids is None:
            batch_size = encoder_attention_mask.shape[0]
            decoder_input_ids = (
                torch.ones(
                    (batch_size, 1),
                    dtype=torch.long,
                    device=encoder_attention_mask.device,
                )
                * self.decoder_start_token_id
            )

        decoder_outputs = self.decoder(
            input_ids=decoder_input_ids,
            encoder_hidden_states=encoder_hidden_states,
            encoder_attention_mask=encoder_attention_mask,
            use_cache=True,
            return_dict=True,
        )

        sequence_output = decoder_outputs.last_hidden_state
        present_key_values = decoder_outputs.past_key_values

        if self.tie_word_embeddings:
            sequence_output = sequence_output * (self.config.d_model**-0.5)

        lm_logits = self.lm_head(sequence_output)
        past_self, past_cross = PastKeyValuesHelper.group_by_self_or_cross(present_key_values)
        return lm_logits, past_self, past_cross


class T5Decoder(torch.nn.Module):
    """A T5 decoder with LM head and past key values"""

    def __init__(self, decoder, lm_head, config):
        super().__init__()
        self.decoder = decoder
        self.lm_head = lm_head
        self.config = config
        self.tie_word_embeddings = (
            self.config.tie_word_embeddings if hasattr(self.config, "tie_word_embeddings") else True
        )

    def forward(self, decoder_input_ids, encoder_attention_mask, *past):
        num_decoder_layers = self.config.num_decoder_layers
        past_key_values = PastKeyValuesHelper.group_by_layer(past, num_decoder_layers)

        # This is a hack since only the third dimension of encoder_hidden_states is used here
        dummy_encoder_hidden_states = encoder_attention_mask.unsqueeze(2)
        decoder_outputs = self.decoder(
            input_ids=decoder_input_ids,
            past_key_values=past_key_values,
            encoder_hidden_states=dummy_encoder_hidden_states,
            encoder_attention_mask=encoder_attention_mask,
            use_cache=True,
            return_dict=True,
        )

        sequence_output = decoder_outputs.last_hidden_state
        present_key_values = decoder_outputs.past_key_values

        if self.tie_word_embeddings:
            sequence_output = sequence_output * (self.config.d_model**-0.5)

        lm_logits = self.lm_head(sequence_output)
        present_self, _ = PastKeyValuesHelper.group_by_self_or_cross(present_key_values)

        # Do not return present_cross since they are identical to corresponding past_cross input
        return lm_logits, present_self


class T5DecoderInputs:
    def __init__(
        self,
        decoder_input_ids,
        encoder_attention_mask,
        past_key_values=None,
    ):
        self.decoder_input_ids: torch.LongTensor = decoder_input_ids
        self.encoder_attention_mask: torch.LongTensor = encoder_attention_mask
        self.past_key_values: list[torch.FloatTensor] | list[torch.HalfTensor] | None = past_key_values

    @staticmethod
    def create_dummy(
        config: T5Config | MT5Config,
        batch_size: int,
        encode_sequence_length: int,
        past_decode_sequence_length: int,
        device: torch.device,
        float16: bool = False,
        use_int32_inputs: bool = False,
    ):  # -> T5DecoderInputs:
        """Create dummy inputs for T5Decoder.

        Args:
            decoder: decoder
            batch_size (int): batch size
            encode_sequence_length (int): sequence length of input_ids for encoder
            past_decode_sequence_length (int): past sequence length of input_ids for decoder
            device (torch.device): device of output tensors
            float16 (bool): whether the model uses float32 or float16 in input
            use_int32_inputs(bool): whether use int32 instead of int64 for some inputs

        Returns:
            T5DecoderInputs: dummy inputs for decoder
        """
        num_attention_heads: int = config.num_heads
        num_layers: int = config.num_decoder_layers
        vocab_size: int = config.vocab_size

        # Do not use head_size = hidden_size / num_attention_heads here.
        # For example, mt5-small, d_model=512 and num_heads=6
        head_size: int = config.d_kv

        sequence_length: int = 1  # fixed for decoding
        decoder_input_ids = torch.randint(
            low=0,
            high=vocab_size - 1,
            size=(batch_size, sequence_length),
            dtype=(torch.int32 if use_int32_inputs else torch.int64),
            device=device,
        )

        encoder_inputs = T5EncoderInputs.create_dummy(
            batch_size,
            encode_sequence_length,
            vocab_size,
            device,
            use_int32_inputs=use_int32_inputs,
        )

        float_type = torch.float16 if float16 else torch.float32

        if past_decode_sequence_length > 0:
            self_attention_past_shape = [
                batch_size,
                num_attention_heads,
                past_decode_sequence_length,
                head_size,
            ]
            cross_attention_past_shape = [
                batch_size,
                num_attention_heads,
                encode_sequence_length,
                head_size,
            ]

            past = []
            for _ in range(2 * num_layers):
                past.append(torch.rand(self_attention_past_shape, dtype=float_type, device=device))

            for _ in range(2 * num_layers):
                past.append(torch.rand(cross_attention_past_shape, dtype=float_type, device=device))
        else:
            past = None

        return T5DecoderInputs(decoder_input_ids, encoder_inputs.attention_mask, past)

    def to_list(self) -> list:
        input_list = [
            self.decoder_input_ids,
            self.encoder_attention_mask,
        ]
        if self.past_key_values:
            input_list.extend(self.past_key_values)
        return input_list

    def to_fp32(self):
        past = [p.to(dtype=torch.float32) for p in self.past_key_values] if self.past_key_values else None
        return T5DecoderInputs(
            self.decoder_input_ids.clone(),
            self.encoder_attention_mask.clone(),
            past,
        )


class T5DecoderHelper:
    @staticmethod
    def export_onnx(
        decoder: T5Decoder | T5DecoderInit,
        device: torch.device,
        onnx_model_path: str,
        verbose: bool = True,
        use_external_data_format: bool = False,
        use_int32_inputs: bool = False,
    ):
        """Export decoder to ONNX

        Args:
            decoder (Union[T5Decoder, T5DecoderNoPastState]): decoder object
            device (torch.device): device of decoder object
            onnx_model_path (str): onnx path
            verbose (bool, optional): print verbose information. Defaults to True.
            use_external_data_format (bool, optional): use external data format or not. Defaults to False.
            use_int32_inputs (bool, optional): use int32 inputs
        """
        assert isinstance(decoder, (T5Decoder, T5DecoderInit))

        inputs = T5DecoderInputs.create_dummy(
            decoder.config,
            batch_size=2,
            encode_sequence_length=3,
            past_decode_sequence_length=5 if isinstance(decoder, T5Decoder) else 0,
            device=device,
            use_int32_inputs=use_int32_inputs,
        )
        input_list = inputs.to_list()

        num_decoder_layers = decoder.config.num_decoder_layers

        past_names = PastKeyValuesHelper.get_past_names(num_decoder_layers, present=False)
        present_names = PastKeyValuesHelper.get_past_names(num_decoder_layers, present=True)
        present_self_names = present_names[: 2 * num_decoder_layers]

        input_past_names = past_names if isinstance(decoder, T5Decoder) else []
        output_present_names = present_self_names if isinstance(decoder, T5Decoder) else present_names
        output_names = ["logits", *output_present_names]

        # Shape of input tensors (sequence_length==1):
        #    input_ids: (batch_size, sequence_length)
        #    encoder_attention_mask: (batch_size, encode_sequence_length)
        #    past_self_*: (batch_size, num_heads, past_decode_sequence_length, head_size)
        #    past_cross_*: (batch_size, num_heads, encode_sequence_length, head_size)

        # Shape of output tensors:
        #    logits: (batch_size, sequence_length, vocab_size)
        #    past_self_*: (batch_size, num_heads, past_decode_sequence_length + sequence_length, head_size)
        #    past_cross_*: (batch_size, num_heads, encode_sequence_length, head_size)

        input_names = ["input_ids"]
        input_names.append("encoder_attention_mask")
        input_names.extend(input_past_names)

        dynamic_axes = {
            "input_ids": {
                0: "batch_size",
                # 1: 'sequence_length'
            },
            "encoder_attention_mask": {0: "batch_size", 1: "encode_sequence_length"},
            "encoder_hidden_states": {0: "batch_size", 1: "encode_sequence_length"},
            "logits": {
                0: "batch_size",
                # 1: 'sequence_length'
            },
        }

        for name in input_past_names:
            dynamic_axes[name] = {
                0: "batch_size",
                2: "past_decode_sequence_length" if "self" in name else "encode_sequence_length",
            }

        for name in output_present_names:
            if "cross" in name:
                dynamic_axes[name] = {0: "batch_size", 2: "encode_sequence_length"}
            else:  # self attention past state
                if isinstance(decoder, T5Decoder):
                    dynamic_axes[name] = {
                        0: "batch_size",
                        2: "past_decode_sequence_length + 1",
                    }
                else:
                    dynamic_axes[name] = {
                        0: "batch_size",
                        # 2: 'sequence_length'
                    }

        Path(onnx_model_path).parent.mkdir(parents=True, exist_ok=True)

        with tempfile.TemporaryDirectory() as tmp_dir_name:
            temp_onnx_model_path = os.path.join(tmp_dir_name, "decoder.onnx")
            Path(temp_onnx_model_path).parent.mkdir(parents=True, exist_ok=True)
            torch_onnx_export(
                decoder,
                args=tuple(input_list),
                f=temp_onnx_model_path if use_external_data_format else onnx_model_path,
                export_params=True,
                input_names=input_names,
                output_names=output_names,
                dynamic_axes=dynamic_axes,
                opset_version=12,
                do_constant_folding=True,
                use_external_data_format=use_external_data_format,
                verbose=verbose,
            )

            if use_external_data_format:
                model = onnx.load_model(temp_onnx_model_path, load_external_data=True)
                OnnxModel.save(
                    model,
                    onnx_model_path,
                    save_as_external_data=True,
                    all_tensors_to_one_file=True,
                )

    @staticmethod
    def onnxruntime_inference(ort_session, inputs: T5DecoderInputs):
        """Run inference of ONNX model."""
        logger.debug("start onnxruntime_inference")

        ort_inputs = {
            "input_ids": numpy.ascontiguousarray(inputs.decoder_input_ids.cpu().numpy()),
            "encoder_attention_mask": numpy.ascontiguousarray(inputs.encoder_attention_mask.cpu().numpy()),
        }

        if inputs.past_key_values:
            assert len(inputs.past_key_values) % 4 == 0
            num_layers = int(len(inputs.past_key_values) / 4)
            past_names = PastKeyValuesHelper.get_past_names(num_layers)
            for i, past_tensor in enumerate(inputs.past_key_values):
                ort_inputs[past_names[i]] = numpy.ascontiguousarray(past_tensor.cpu().numpy())

        ort_outputs = ort_session.run(None, ort_inputs)
        return ort_outputs

    @staticmethod
    def verify_onnx(
        model: T5Decoder | T5DecoderInit,
        ort_session: InferenceSession,
        device: torch.device,
        use_int32_inputs: bool,
        max_cases: int = 4,
    ):
        """Compare the result from PyTorch and OnnxRuntime to verify the ONNX model is good."""
        float16: bool = TypeHelper.get_input_type(ort_session, "past_key_self_0") == "tensor(float16)"

        test_cases = [(4, 11, 3), (1, 2, 5), (3, 1, 1), (8, 5, 2)]
        test_cases_max_diff = []
        for (
            batch_size,
            encode_sequence_length,
            past_decode_sequence_length,
        ) in test_cases[:max_cases]:
            if isinstance(model, T5DecoderInit):
                past_decode_sequence_length = 0  # noqa: PLW2901

            inputs = T5DecoderInputs.create_dummy(
                model.config,
                batch_size,
                encode_sequence_length,
                past_decode_sequence_length,
                device=device,
                float16=float16,
                use_int32_inputs=use_int32_inputs,
            )

            # We use fp32 PyTroch model as baseline even when ONNX model is fp16
            input_list = inputs.to_fp32().to_list()

            # Run inference of PyTorch model
            with torch.no_grad():
                torch_outputs = model(*input_list)

            ort_outputs = T5DecoderHelper.onnxruntime_inference(ort_session, inputs)
            num_decoder_layers = model.config.num_decoder_layers

            max_diff = numpy.amax(numpy.abs(torch_outputs[0].cpu().numpy() - ort_outputs[0]))
            max_diff_all = max_diff
            logger.debug(f"logits max_diff={max_diff}")

            for i in range(2 * num_decoder_layers):
                max_diff = numpy.amax(numpy.abs(torch_outputs[1][i].cpu().numpy() - ort_outputs[1 + i]))
                logger.debug(f"self attention past state {i} max_diff={max_diff}")
                max_diff_all = max(max_diff_all, max_diff)

            if isinstance(model, T5DecoderInit):
                for i in range(2 * num_decoder_layers):
                    max_diff = numpy.amax(
                        numpy.abs(torch_outputs[2][i].cpu().numpy() - ort_outputs[1 + 2 * num_decoder_layers + i])
                    )
                    logger.debug(f"cross attention past state {i} max_diff={max_diff}")
                    max_diff_all = max(max_diff_all, max_diff)

            test_cases_max_diff.append(max_diff_all)
            logger.info(
                "batch_size=%s, encode_sequence_length=%s, past_decode_sequence_length=%s, max_diff=%s",
                batch_size,
                encode_sequence_length,
                past_decode_sequence_length,
                max_diff_all,
            )

        return max_diff_all

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\dtypes\inference.py ===
""" basic inference routines """

from __future__ import annotations

from collections import abc
from numbers import Number
import re
from re import Pattern
from typing import TYPE_CHECKING

import numpy as np

from pandas._libs import lib

if TYPE_CHECKING:
    from collections.abc import Hashable

    from pandas._typing import TypeGuard

is_bool = lib.is_bool

is_integer = lib.is_integer

is_float = lib.is_float

is_complex = lib.is_complex

is_scalar = lib.is_scalar

is_decimal = lib.is_decimal

is_interval = lib.is_interval

is_list_like = lib.is_list_like

is_iterator = lib.is_iterator


def is_number(obj) -> TypeGuard[Number | np.number]:
    """
    Check if the object is a number.

    Returns True when the object is a number, and False if is not.

    Parameters
    ----------
    obj : any type
        The object to check if is a number.

    Returns
    -------
    bool
        Whether `obj` is a number or not.

    See Also
    --------
    api.types.is_integer: Checks a subgroup of numbers.

    Examples
    --------
    >>> from pandas.api.types import is_number
    >>> is_number(1)
    True
    >>> is_number(7.15)
    True

    Booleans are valid because they are int subclass.

    >>> is_number(False)
    True

    >>> is_number("foo")
    False
    >>> is_number("5")
    False
    """
    return isinstance(obj, (Number, np.number))


def iterable_not_string(obj) -> bool:
    """
    Check if the object is an iterable but not a string.

    Parameters
    ----------
    obj : The object to check.

    Returns
    -------
    is_iter_not_string : bool
        Whether `obj` is a non-string iterable.

    Examples
    --------
    >>> iterable_not_string([1, 2, 3])
    True
    >>> iterable_not_string("foo")
    False
    >>> iterable_not_string(1)
    False
    """
    return isinstance(obj, abc.Iterable) and not isinstance(obj, str)


def is_file_like(obj) -> bool:
    """
    Check if the object is a file-like object.

    For objects to be considered file-like, they must
    be an iterator AND have either a `read` and/or `write`
    method as an attribute.

    Note: file-like objects must be iterable, but
    iterable objects need not be file-like.

    Parameters
    ----------
    obj : The object to check

    Returns
    -------
    bool
        Whether `obj` has file-like properties.

    Examples
    --------
    >>> import io
    >>> from pandas.api.types import is_file_like
    >>> buffer = io.StringIO("data")
    >>> is_file_like(buffer)
    True
    >>> is_file_like([1, 2, 3])
    False
    """
    if not (hasattr(obj, "read") or hasattr(obj, "write")):
        return False

    return bool(hasattr(obj, "__iter__"))


def is_re(obj) -> TypeGuard[Pattern]:
    """
    Check if the object is a regex pattern instance.

    Parameters
    ----------
    obj : The object to check

    Returns
    -------
    bool
        Whether `obj` is a regex pattern.

    Examples
    --------
    >>> from pandas.api.types import is_re
    >>> import re
    >>> is_re(re.compile(".*"))
    True
    >>> is_re("foo")
    False
    """
    return isinstance(obj, Pattern)


def is_re_compilable(obj) -> bool:
    """
    Check if the object can be compiled into a regex pattern instance.

    Parameters
    ----------
    obj : The object to check

    Returns
    -------
    bool
        Whether `obj` can be compiled as a regex pattern.

    Examples
    --------
    >>> from pandas.api.types import is_re_compilable
    >>> is_re_compilable(".*")
    True
    >>> is_re_compilable(1)
    False
    """
    try:
        re.compile(obj)
    except TypeError:
        return False
    else:
        return True


def is_array_like(obj) -> bool:
    """
    Check if the object is array-like.

    For an object to be considered array-like, it must be list-like and
    have a `dtype` attribute.

    Parameters
    ----------
    obj : The object to check

    Returns
    -------
    is_array_like : bool
        Whether `obj` has array-like properties.

    Examples
    --------
    >>> is_array_like(np.array([1, 2, 3]))
    True
    >>> is_array_like(pd.Series(["a", "b"]))
    True
    >>> is_array_like(pd.Index(["2016-01-01"]))
    True
    >>> is_array_like([1, 2, 3])
    False
    >>> is_array_like(("a", "b"))
    False
    """
    return is_list_like(obj) and hasattr(obj, "dtype")


def is_nested_list_like(obj) -> bool:
    """
    Check if the object is list-like, and that all of its elements
    are also list-like.

    Parameters
    ----------
    obj : The object to check

    Returns
    -------
    is_list_like : bool
        Whether `obj` has list-like properties.

    Examples
    --------
    >>> is_nested_list_like([[1, 2, 3]])
    True
    >>> is_nested_list_like([{1, 2, 3}, {1, 2, 3}])
    True
    >>> is_nested_list_like(["foo"])
    False
    >>> is_nested_list_like([])
    False
    >>> is_nested_list_like([[1, 2, 3], 1])
    False

    Notes
    -----
    This won't reliably detect whether a consumable iterator (e. g.
    a generator) is a nested-list-like without consuming the iterator.
    To avoid consuming it, we always return False if the outer container
    doesn't define `__len__`.

    See Also
    --------
    is_list_like
    """
    return (
        is_list_like(obj)
        and hasattr(obj, "__len__")
        and len(obj) > 0
        and all(is_list_like(item) for item in obj)
    )


def is_dict_like(obj) -> bool:
    """
    Check if the object is dict-like.

    Parameters
    ----------
    obj : The object to check

    Returns
    -------
    bool
        Whether `obj` has dict-like properties.

    Examples
    --------
    >>> from pandas.api.types import is_dict_like
    >>> is_dict_like({1: 2})
    True
    >>> is_dict_like([1, 2, 3])
    False
    >>> is_dict_like(dict)
    False
    >>> is_dict_like(dict())
    True
    """
    dict_like_attrs = ("__getitem__", "keys", "__contains__")
    return (
        all(hasattr(obj, attr) for attr in dict_like_attrs)
        # [GH 25196] exclude classes
        and not isinstance(obj, type)
    )


def is_named_tuple(obj) -> bool:
    """
    Check if the object is a named tuple.

    Parameters
    ----------
    obj : The object to check

    Returns
    -------
    bool
        Whether `obj` is a named tuple.

    Examples
    --------
    >>> from collections import namedtuple
    >>> from pandas.api.types import is_named_tuple
    >>> Point = namedtuple("Point", ["x", "y"])
    >>> p = Point(1, 2)
    >>>
    >>> is_named_tuple(p)
    True
    >>> is_named_tuple((1, 2))
    False
    """
    return isinstance(obj, abc.Sequence) and hasattr(obj, "_fields")


def is_hashable(obj) -> TypeGuard[Hashable]:
    """
    Return True if hash(obj) will succeed, False otherwise.

    Some types will pass a test against collections.abc.Hashable but fail when
    they are actually hashed with hash().

    Distinguish between these and other types by trying the call to hash() and
    seeing if they raise TypeError.

    Returns
    -------
    bool

    Examples
    --------
    >>> import collections
    >>> from pandas.api.types import is_hashable
    >>> a = ([],)
    >>> isinstance(a, collections.abc.Hashable)
    True
    >>> is_hashable(a)
    False
    """
    # Unfortunately, we can't use isinstance(obj, collections.abc.Hashable),
    # which can be faster than calling hash. That is because numpy scalars
    # fail this test.

    # Reconsider this decision once this numpy bug is fixed:
    # https://github.com/numpy/numpy/issues/5562

    try:
        hash(obj)
    except TypeError:
        return False
    else:
        return True


def is_sequence(obj) -> bool:
    """
    Check if the object is a sequence of objects.
    String types are not included as sequences here.

    Parameters
    ----------
    obj : The object to check

    Returns
    -------
    is_sequence : bool
        Whether `obj` is a sequence of objects.

    Examples
    --------
    >>> l = [1, 2, 3]
    >>>
    >>> is_sequence(l)
    True
    >>> is_sequence(iter(l))
    False
    """
    try:
        iter(obj)  # Can iterate over it.
        len(obj)  # Has a length associated with it.
        return not isinstance(obj, (str, bytes))
    except (TypeError, AttributeError):
        return False


def is_dataclass(item) -> bool:
    """
    Checks if the object is a data-class instance

    Parameters
    ----------
    item : object

    Returns
    --------
    is_dataclass : bool
        True if the item is an instance of a data-class,
        will return false if you pass the data class itself

    Examples
    --------
    >>> from dataclasses import dataclass
    >>> @dataclass
    ... class Point:
    ...     x: int
    ...     y: int

    >>> is_dataclass(Point)
    False
    >>> is_dataclass(Point(0,2))
    True

    """
    try:
        import dataclasses

        return dataclasses.is_dataclass(item) and not isinstance(item, type)
    except ImportError:
        return False

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\idna\core.py ===
import bisect
import re
import unicodedata
from typing import Optional, Union

from . import idnadata
from .intranges import intranges_contain

_virama_combining_class = 9
_alabel_prefix = b"xn--"
_unicode_dots_re = re.compile("[\u002e\u3002\uff0e\uff61]")


class IDNAError(UnicodeError):
    """Base exception for all IDNA-encoding related problems"""

    pass


class IDNABidiError(IDNAError):
    """Exception when bidirectional requirements are not satisfied"""

    pass


class InvalidCodepoint(IDNAError):
    """Exception when a disallowed or unallocated codepoint is used"""

    pass


class InvalidCodepointContext(IDNAError):
    """Exception when the codepoint is not valid in the context it is used"""

    pass


def _combining_class(cp: int) -> int:
    v = unicodedata.combining(chr(cp))
    if v == 0:
        if not unicodedata.name(chr(cp)):
            raise ValueError("Unknown character in unicodedata")
    return v


def _is_script(cp: str, script: str) -> bool:
    return intranges_contain(ord(cp), idnadata.scripts[script])


def _punycode(s: str) -> bytes:
    return s.encode("punycode")


def _unot(s: int) -> str:
    return "U+{:04X}".format(s)


def valid_label_length(label: Union[bytes, str]) -> bool:
    if len(label) > 63:
        return False
    return True


def valid_string_length(label: Union[bytes, str], trailing_dot: bool) -> bool:
    if len(label) > (254 if trailing_dot else 253):
        return False
    return True


def check_bidi(label: str, check_ltr: bool = False) -> bool:
    # Bidi rules should only be applied if string contains RTL characters
    bidi_label = False
    for idx, cp in enumerate(label, 1):
        direction = unicodedata.bidirectional(cp)
        if direction == "":
            # String likely comes from a newer version of Unicode
            raise IDNABidiError("Unknown directionality in label {} at position {}".format(repr(label), idx))
        if direction in ["R", "AL", "AN"]:
            bidi_label = True
    if not bidi_label and not check_ltr:
        return True

    # Bidi rule 1
    direction = unicodedata.bidirectional(label[0])
    if direction in ["R", "AL"]:
        rtl = True
    elif direction == "L":
        rtl = False
    else:
        raise IDNABidiError("First codepoint in label {} must be directionality L, R or AL".format(repr(label)))

    valid_ending = False
    number_type: Optional[str] = None
    for idx, cp in enumerate(label, 1):
        direction = unicodedata.bidirectional(cp)

        if rtl:
            # Bidi rule 2
            if direction not in [
                "R",
                "AL",
                "AN",
                "EN",
                "ES",
                "CS",
                "ET",
                "ON",
                "BN",
                "NSM",
            ]:
                raise IDNABidiError("Invalid direction for codepoint at position {} in a right-to-left label".format(idx))
            # Bidi rule 3
            if direction in ["R", "AL", "EN", "AN"]:
                valid_ending = True
            elif direction != "NSM":
                valid_ending = False
            # Bidi rule 4
            if direction in ["AN", "EN"]:
                if not number_type:
                    number_type = direction
                else:
                    if number_type != direction:
                        raise IDNABidiError("Can not mix numeral types in a right-to-left label")
        else:
            # Bidi rule 5
            if direction not in ["L", "EN", "ES", "CS", "ET", "ON", "BN", "NSM"]:
                raise IDNABidiError("Invalid direction for codepoint at position {} in a left-to-right label".format(idx))
            # Bidi rule 6
            if direction in ["L", "EN"]:
                valid_ending = True
            elif direction != "NSM":
                valid_ending = False

    if not valid_ending:
        raise IDNABidiError("Label ends with illegal codepoint directionality")

    return True


def check_initial_combiner(label: str) -> bool:
    if unicodedata.category(label[0])[0] == "M":
        raise IDNAError("Label begins with an illegal combining character")
    return True


def check_hyphen_ok(label: str) -> bool:
    if label[2:4] == "--":
        raise IDNAError("Label has disallowed hyphens in 3rd and 4th position")
    if label[0] == "-" or label[-1] == "-":
        raise IDNAError("Label must not start or end with a hyphen")
    return True


def check_nfc(label: str) -> None:
    if unicodedata.normalize("NFC", label) != label:
        raise IDNAError("Label must be in Normalization Form C")


def valid_contextj(label: str, pos: int) -> bool:
    cp_value = ord(label[pos])

    if cp_value == 0x200C:
        if pos > 0:
            if _combining_class(ord(label[pos - 1])) == _virama_combining_class:
                return True

        ok = False
        for i in range(pos - 1, -1, -1):
            joining_type = idnadata.joining_types.get(ord(label[i]))
            if joining_type == ord("T"):
                continue
            elif joining_type in [ord("L"), ord("D")]:
                ok = True
                break
            else:
                break

        if not ok:
            return False

        ok = False
        for i in range(pos + 1, len(label)):
            joining_type = idnadata.joining_types.get(ord(label[i]))
            if joining_type == ord("T"):
                continue
            elif joining_type in [ord("R"), ord("D")]:
                ok = True
                break
            else:
                break
        return ok

    if cp_value == 0x200D:
        if pos > 0:
            if _combining_class(ord(label[pos - 1])) == _virama_combining_class:
                return True
        return False

    else:
        return False


def valid_contexto(label: str, pos: int, exception: bool = False) -> bool:
    cp_value = ord(label[pos])

    if cp_value == 0x00B7:
        if 0 < pos < len(label) - 1:
            if ord(label[pos - 1]) == 0x006C and ord(label[pos + 1]) == 0x006C:
                return True
        return False

    elif cp_value == 0x0375:
        if pos < len(label) - 1 and len(label) > 1:
            return _is_script(label[pos + 1], "Greek")
        return False

    elif cp_value == 0x05F3 or cp_value == 0x05F4:
        if pos > 0:
            return _is_script(label[pos - 1], "Hebrew")
        return False

    elif cp_value == 0x30FB:
        for cp in label:
            if cp == "\u30fb":
                continue
            if _is_script(cp, "Hiragana") or _is_script(cp, "Katakana") or _is_script(cp, "Han"):
                return True
        return False

    elif 0x660 <= cp_value <= 0x669:
        for cp in label:
            if 0x6F0 <= ord(cp) <= 0x06F9:
                return False
        return True

    elif 0x6F0 <= cp_value <= 0x6F9:
        for cp in label:
            if 0x660 <= ord(cp) <= 0x0669:
                return False
        return True

    return False


def check_label(label: Union[str, bytes, bytearray]) -> None:
    if isinstance(label, (bytes, bytearray)):
        label = label.decode("utf-8")
    if len(label) == 0:
        raise IDNAError("Empty Label")

    check_nfc(label)
    check_hyphen_ok(label)
    check_initial_combiner(label)

    for pos, cp in enumerate(label):
        cp_value = ord(cp)
        if intranges_contain(cp_value, idnadata.codepoint_classes["PVALID"]):
            continue
        elif intranges_contain(cp_value, idnadata.codepoint_classes["CONTEXTJ"]):
            try:
                if not valid_contextj(label, pos):
                    raise InvalidCodepointContext(
                        "Joiner {} not allowed at position {} in {}".format(_unot(cp_value), pos + 1, repr(label))
                    )
            except ValueError:
                raise IDNAError(
                    "Unknown codepoint adjacent to joiner {} at position {} in {}".format(
                        _unot(cp_value), pos + 1, repr(label)
                    )
                )
        elif intranges_contain(cp_value, idnadata.codepoint_classes["CONTEXTO"]):
            if not valid_contexto(label, pos):
                raise InvalidCodepointContext(
                    "Codepoint {} not allowed at position {} in {}".format(_unot(cp_value), pos + 1, repr(label))
                )
        else:
            raise InvalidCodepoint(
                "Codepoint {} at position {} of {} not allowed".format(_unot(cp_value), pos + 1, repr(label))
            )

    check_bidi(label)


def alabel(label: str) -> bytes:
    try:
        label_bytes = label.encode("ascii")
        ulabel(label_bytes)
        if not valid_label_length(label_bytes):
            raise IDNAError("Label too long")
        return label_bytes
    except UnicodeEncodeError:
        pass

    check_label(label)
    label_bytes = _alabel_prefix + _punycode(label)

    if not valid_label_length(label_bytes):
        raise IDNAError("Label too long")

    return label_bytes


def ulabel(label: Union[str, bytes, bytearray]) -> str:
    if not isinstance(label, (bytes, bytearray)):
        try:
            label_bytes = label.encode("ascii")
        except UnicodeEncodeError:
            check_label(label)
            return label
    else:
        label_bytes = label

    label_bytes = label_bytes.lower()
    if label_bytes.startswith(_alabel_prefix):
        label_bytes = label_bytes[len(_alabel_prefix) :]
        if not label_bytes:
            raise IDNAError("Malformed A-label, no Punycode eligible content found")
        if label_bytes.decode("ascii")[-1] == "-":
            raise IDNAError("A-label must not end with a hyphen")
    else:
        check_label(label_bytes)
        return label_bytes.decode("ascii")

    try:
        label = label_bytes.decode("punycode")
    except UnicodeError:
        raise IDNAError("Invalid A-label")
    check_label(label)
    return label


def uts46_remap(domain: str, std3_rules: bool = True, transitional: bool = False) -> str:
    """Re-map the characters in the string according to UTS46 processing."""
    from .uts46data import uts46data

    output = ""

    for pos, char in enumerate(domain):
        code_point = ord(char)
        try:
            uts46row = uts46data[code_point if code_point < 256 else bisect.bisect_left(uts46data, (code_point, "Z")) - 1]
            status = uts46row[1]
            replacement: Optional[str] = None
            if len(uts46row) == 3:
                replacement = uts46row[2]
            if (
                status == "V"
                or (status == "D" and not transitional)
                or (status == "3" and not std3_rules and replacement is None)
            ):
                output += char
            elif replacement is not None and (
                status == "M" or (status == "3" and not std3_rules) or (status == "D" and transitional)
            ):
                output += replacement
            elif status != "I":
                raise IndexError()
        except IndexError:
            raise InvalidCodepoint(
                "Codepoint {} not allowed at position {} in {}".format(_unot(code_point), pos + 1, repr(domain))
            )

    return unicodedata.normalize("NFC", output)


def encode(
    s: Union[str, bytes, bytearray],
    strict: bool = False,
    uts46: bool = False,
    std3_rules: bool = False,
    transitional: bool = False,
) -> bytes:
    if not isinstance(s, str):
        try:
            s = str(s, "ascii")
        except UnicodeDecodeError:
            raise IDNAError("should pass a unicode string to the function rather than a byte string.")
    if uts46:
        s = uts46_remap(s, std3_rules, transitional)
    trailing_dot = False
    result = []
    if strict:
        labels = s.split(".")
    else:
        labels = _unicode_dots_re.split(s)
    if not labels or labels == [""]:
        raise IDNAError("Empty domain")
    if labels[-1] == "":
        del labels[-1]
        trailing_dot = True
    for label in labels:
        s = alabel(label)
        if s:
            result.append(s)
        else:
            raise IDNAError("Empty label")
    if trailing_dot:
        result.append(b"")
    s = b".".join(result)
    if not valid_string_length(s, trailing_dot):
        raise IDNAError("Domain too long")
    return s


def decode(
    s: Union[str, bytes, bytearray],
    strict: bool = False,
    uts46: bool = False,
    std3_rules: bool = False,
) -> str:
    try:
        if not isinstance(s, str):
            s = str(s, "ascii")
    except UnicodeDecodeError:
        raise IDNAError("Invalid ASCII in A-label")
    if uts46:
        s = uts46_remap(s, std3_rules, False)
    trailing_dot = False
    result = []
    if not strict:
        labels = _unicode_dots_re.split(s)
    else:
        labels = s.split(".")
    if not labels or labels == [""]:
        raise IDNAError("Empty domain")
    if not labels[-1]:
        del labels[-1]
        trailing_dot = True
    for label in labels:
        s = ulabel(label)
        if s:
            result.append(s)
        else:
            raise IDNAError("Empty label")
    if trailing_dot:
        result.append("")
    return ".".join(result)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\jinja2\sandbox.py ===
"""A sandbox layer that ensures unsafe operations cannot be performed.
Useful when the template itself comes from an untrusted source.
"""

import operator
import types
import typing as t
from _string import formatter_field_name_split  # type: ignore
from collections import abc
from collections import deque
from functools import update_wrapper
from string import Formatter

from markupsafe import EscapeFormatter
from markupsafe import Markup

from .environment import Environment
from .exceptions import SecurityError
from .runtime import Context
from .runtime import Undefined

F = t.TypeVar("F", bound=t.Callable[..., t.Any])

#: maximum number of items a range may produce
MAX_RANGE = 100000

#: Unsafe function attributes.
UNSAFE_FUNCTION_ATTRIBUTES: t.Set[str] = set()

#: Unsafe method attributes. Function attributes are unsafe for methods too.
UNSAFE_METHOD_ATTRIBUTES: t.Set[str] = set()

#: unsafe generator attributes.
UNSAFE_GENERATOR_ATTRIBUTES = {"gi_frame", "gi_code"}

#: unsafe attributes on coroutines
UNSAFE_COROUTINE_ATTRIBUTES = {"cr_frame", "cr_code"}

#: unsafe attributes on async generators
UNSAFE_ASYNC_GENERATOR_ATTRIBUTES = {"ag_code", "ag_frame"}

_mutable_spec: t.Tuple[t.Tuple[t.Type[t.Any], t.FrozenSet[str]], ...] = (
    (
        abc.MutableSet,
        frozenset(
            [
                "add",
                "clear",
                "difference_update",
                "discard",
                "pop",
                "remove",
                "symmetric_difference_update",
                "update",
            ]
        ),
    ),
    (
        abc.MutableMapping,
        frozenset(["clear", "pop", "popitem", "setdefault", "update"]),
    ),
    (
        abc.MutableSequence,
        frozenset(
            ["append", "clear", "pop", "reverse", "insert", "sort", "extend", "remove"]
        ),
    ),
    (
        deque,
        frozenset(
            [
                "append",
                "appendleft",
                "clear",
                "extend",
                "extendleft",
                "pop",
                "popleft",
                "remove",
                "rotate",
            ]
        ),
    ),
)


def safe_range(*args: int) -> range:
    """A range that can't generate ranges with a length of more than
    MAX_RANGE items.
    """
    rng = range(*args)

    if len(rng) > MAX_RANGE:
        raise OverflowError(
            "Range too big. The sandbox blocks ranges larger than"
            f" MAX_RANGE ({MAX_RANGE})."
        )

    return rng


def unsafe(f: F) -> F:
    """Marks a function or method as unsafe.

    .. code-block: python

        @unsafe
        def delete(self):
            pass
    """
    f.unsafe_callable = True  # type: ignore
    return f


def is_internal_attribute(obj: t.Any, attr: str) -> bool:
    """Test if the attribute given is an internal python attribute.  For
    example this function returns `True` for the `func_code` attribute of
    python objects.  This is useful if the environment method
    :meth:`~SandboxedEnvironment.is_safe_attribute` is overridden.

    >>> from jinja2.sandbox import is_internal_attribute
    >>> is_internal_attribute(str, "mro")
    True
    >>> is_internal_attribute(str, "upper")
    False
    """
    if isinstance(obj, types.FunctionType):
        if attr in UNSAFE_FUNCTION_ATTRIBUTES:
            return True
    elif isinstance(obj, types.MethodType):
        if attr in UNSAFE_FUNCTION_ATTRIBUTES or attr in UNSAFE_METHOD_ATTRIBUTES:
            return True
    elif isinstance(obj, type):
        if attr == "mro":
            return True
    elif isinstance(obj, (types.CodeType, types.TracebackType, types.FrameType)):
        return True
    elif isinstance(obj, types.GeneratorType):
        if attr in UNSAFE_GENERATOR_ATTRIBUTES:
            return True
    elif hasattr(types, "CoroutineType") and isinstance(obj, types.CoroutineType):
        if attr in UNSAFE_COROUTINE_ATTRIBUTES:
            return True
    elif hasattr(types, "AsyncGeneratorType") and isinstance(
        obj, types.AsyncGeneratorType
    ):
        if attr in UNSAFE_ASYNC_GENERATOR_ATTRIBUTES:
            return True
    return attr.startswith("__")


def modifies_known_mutable(obj: t.Any, attr: str) -> bool:
    """This function checks if an attribute on a builtin mutable object
    (list, dict, set or deque) or the corresponding ABCs would modify it
    if called.

    >>> modifies_known_mutable({}, "clear")
    True
    >>> modifies_known_mutable({}, "keys")
    False
    >>> modifies_known_mutable([], "append")
    True
    >>> modifies_known_mutable([], "index")
    False

    If called with an unsupported object, ``False`` is returned.

    >>> modifies_known_mutable("foo", "upper")
    False
    """
    for typespec, unsafe in _mutable_spec:
        if isinstance(obj, typespec):
            return attr in unsafe
    return False


class SandboxedEnvironment(Environment):
    """The sandboxed environment.  It works like the regular environment but
    tells the compiler to generate sandboxed code.  Additionally subclasses of
    this environment may override the methods that tell the runtime what
    attributes or functions are safe to access.

    If the template tries to access insecure code a :exc:`SecurityError` is
    raised.  However also other exceptions may occur during the rendering so
    the caller has to ensure that all exceptions are caught.
    """

    sandboxed = True

    #: default callback table for the binary operators.  A copy of this is
    #: available on each instance of a sandboxed environment as
    #: :attr:`binop_table`
    default_binop_table: t.Dict[str, t.Callable[[t.Any, t.Any], t.Any]] = {
        "+": operator.add,
        "-": operator.sub,
        "*": operator.mul,
        "/": operator.truediv,
        "//": operator.floordiv,
        "**": operator.pow,
        "%": operator.mod,
    }

    #: default callback table for the unary operators.  A copy of this is
    #: available on each instance of a sandboxed environment as
    #: :attr:`unop_table`
    default_unop_table: t.Dict[str, t.Callable[[t.Any], t.Any]] = {
        "+": operator.pos,
        "-": operator.neg,
    }

    #: a set of binary operators that should be intercepted.  Each operator
    #: that is added to this set (empty by default) is delegated to the
    #: :meth:`call_binop` method that will perform the operator.  The default
    #: operator callback is specified by :attr:`binop_table`.
    #:
    #: The following binary operators are interceptable:
    #: ``//``, ``%``, ``+``, ``*``, ``-``, ``/``, and ``**``
    #:
    #: The default operation form the operator table corresponds to the
    #: builtin function.  Intercepted calls are always slower than the native
    #: operator call, so make sure only to intercept the ones you are
    #: interested in.
    #:
    #: .. versionadded:: 2.6
    intercepted_binops: t.FrozenSet[str] = frozenset()

    #: a set of unary operators that should be intercepted.  Each operator
    #: that is added to this set (empty by default) is delegated to the
    #: :meth:`call_unop` method that will perform the operator.  The default
    #: operator callback is specified by :attr:`unop_table`.
    #:
    #: The following unary operators are interceptable: ``+``, ``-``
    #:
    #: The default operation form the operator table corresponds to the
    #: builtin function.  Intercepted calls are always slower than the native
    #: operator call, so make sure only to intercept the ones you are
    #: interested in.
    #:
    #: .. versionadded:: 2.6
    intercepted_unops: t.FrozenSet[str] = frozenset()

    def __init__(self, *args: t.Any, **kwargs: t.Any) -> None:
        super().__init__(*args, **kwargs)
        self.globals["range"] = safe_range
        self.binop_table = self.default_binop_table.copy()
        self.unop_table = self.default_unop_table.copy()

    def is_safe_attribute(self, obj: t.Any, attr: str, value: t.Any) -> bool:
        """The sandboxed environment will call this method to check if the
        attribute of an object is safe to access.  Per default all attributes
        starting with an underscore are considered private as well as the
        special attributes of internal python objects as returned by the
        :func:`is_internal_attribute` function.
        """
        return not (attr.startswith("_") or is_internal_attribute(obj, attr))

    def is_safe_callable(self, obj: t.Any) -> bool:
        """Check if an object is safely callable. By default callables
        are considered safe unless decorated with :func:`unsafe`.

        This also recognizes the Django convention of setting
        ``func.alters_data = True``.
        """
        return not (
            getattr(obj, "unsafe_callable", False) or getattr(obj, "alters_data", False)
        )

    def call_binop(
        self, context: Context, operator: str, left: t.Any, right: t.Any
    ) -> t.Any:
        """For intercepted binary operator calls (:meth:`intercepted_binops`)
        this function is executed instead of the builtin operator.  This can
        be used to fine tune the behavior of certain operators.

        .. versionadded:: 2.6
        """
        return self.binop_table[operator](left, right)

    def call_unop(self, context: Context, operator: str, arg: t.Any) -> t.Any:
        """For intercepted unary operator calls (:meth:`intercepted_unops`)
        this function is executed instead of the builtin operator.  This can
        be used to fine tune the behavior of certain operators.

        .. versionadded:: 2.6
        """
        return self.unop_table[operator](arg)

    def getitem(
        self, obj: t.Any, argument: t.Union[str, t.Any]
    ) -> t.Union[t.Any, Undefined]:
        """Subscribe an object from sandboxed code."""
        try:
            return obj[argument]
        except (TypeError, LookupError):
            if isinstance(argument, str):
                try:
                    attr = str(argument)
                except Exception:
                    pass
                else:
                    try:
                        value = getattr(obj, attr)
                    except AttributeError:
                        pass
                    else:
                        fmt = self.wrap_str_format(value)
                        if fmt is not None:
                            return fmt
                        if self.is_safe_attribute(obj, argument, value):
                            return value
                        return self.unsafe_undefined(obj, argument)
        return self.undefined(obj=obj, name=argument)

    def getattr(self, obj: t.Any, attribute: str) -> t.Union[t.Any, Undefined]:
        """Subscribe an object from sandboxed code and prefer the
        attribute.  The attribute passed *must* be a bytestring.
        """
        try:
            value = getattr(obj, attribute)
        except AttributeError:
            try:
                return obj[attribute]
            except (TypeError, LookupError):
                pass
        else:
            fmt = self.wrap_str_format(value)
            if fmt is not None:
                return fmt
            if self.is_safe_attribute(obj, attribute, value):
                return value
            return self.unsafe_undefined(obj, attribute)
        return self.undefined(obj=obj, name=attribute)

    def unsafe_undefined(self, obj: t.Any, attribute: str) -> Undefined:
        """Return an undefined object for unsafe attributes."""
        return self.undefined(
            f"access to attribute {attribute!r} of"
            f" {type(obj).__name__!r} object is unsafe.",
            name=attribute,
            obj=obj,
            exc=SecurityError,
        )

    def wrap_str_format(self, value: t.Any) -> t.Optional[t.Callable[..., str]]:
        """If the given value is a ``str.format`` or ``str.format_map`` method,
        return a new function than handles sandboxing. This is done at access
        rather than in :meth:`call`, so that calls made without ``call`` are
        also sandboxed.
        """
        if not isinstance(
            value, (types.MethodType, types.BuiltinMethodType)
        ) or value.__name__ not in ("format", "format_map"):
            return None

        f_self: t.Any = value.__self__

        if not isinstance(f_self, str):
            return None

        str_type: t.Type[str] = type(f_self)
        is_format_map = value.__name__ == "format_map"
        formatter: SandboxedFormatter

        if isinstance(f_self, Markup):
            formatter = SandboxedEscapeFormatter(self, escape=f_self.escape)
        else:
            formatter = SandboxedFormatter(self)

        vformat = formatter.vformat

        def wrapper(*args: t.Any, **kwargs: t.Any) -> str:
            if is_format_map:
                if kwargs:
                    raise TypeError("format_map() takes no keyword arguments")

                if len(args) != 1:
                    raise TypeError(
                        f"format_map() takes exactly one argument ({len(args)} given)"
                    )

                kwargs = args[0]
                args = ()

            return str_type(vformat(f_self, args, kwargs))

        return update_wrapper(wrapper, value)

    def call(
        __self,  # noqa: B902
        __context: Context,
        __obj: t.Any,
        *args: t.Any,
        **kwargs: t.Any,
    ) -> t.Any:
        """Call an object from sandboxed code."""

        # the double prefixes are to avoid double keyword argument
        # errors when proxying the call.
        if not __self.is_safe_callable(__obj):
            raise SecurityError(f"{__obj!r} is not safely callable")
        return __context.call(__obj, *args, **kwargs)


class ImmutableSandboxedEnvironment(SandboxedEnvironment):
    """Works exactly like the regular `SandboxedEnvironment` but does not
    permit modifications on the builtin mutable objects `list`, `set`, and
    `dict` by using the :func:`modifies_known_mutable` function.
    """

    def is_safe_attribute(self, obj: t.Any, attr: str, value: t.Any) -> bool:
        if not super().is_safe_attribute(obj, attr, value):
            return False

        return not modifies_known_mutable(obj, attr)


class SandboxedFormatter(Formatter):
    def __init__(self, env: Environment, **kwargs: t.Any) -> None:
        self._env = env
        super().__init__(**kwargs)

    def get_field(
        self, field_name: str, args: t.Sequence[t.Any], kwargs: t.Mapping[str, t.Any]
    ) -> t.Tuple[t.Any, str]:
        first, rest = formatter_field_name_split(field_name)
        obj = self.get_value(first, args, kwargs)
        for is_attr, i in rest:
            if is_attr:
                obj = self._env.getattr(obj, i)
            else:
                obj = self._env.getitem(obj, i)
        return obj, first


class SandboxedEscapeFormatter(SandboxedFormatter, EscapeFormatter):
    pass

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\drawing\colors.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Alias,
    Typed,
    Integer,
    Set,
    MinMax,
)
from openpyxl.descriptors.excel import Percentage
from openpyxl.descriptors.nested import (
    NestedNoneSet,
    NestedValue,
    NestedInteger,
    EmptyTag,
)

from openpyxl.styles.colors import RGB
from openpyxl.xml.constants import DRAWING_NS

from openpyxl.descriptors.excel import ExtensionList as OfficeArtExtensionList

PRESET_COLORS = [
        'aliceBlue', 'antiqueWhite', 'aqua', 'aquamarine',
        'azure', 'beige', 'bisque', 'black', 'blanchedAlmond', 'blue',
        'blueViolet', 'brown', 'burlyWood', 'cadetBlue', 'chartreuse',
        'chocolate', 'coral', 'cornflowerBlue', 'cornsilk', 'crimson', 'cyan',
        'darkBlue', 'darkCyan', 'darkGoldenrod', 'darkGray', 'darkGrey',
        'darkGreen', 'darkKhaki', 'darkMagenta', 'darkOliveGreen', 'darkOrange',
        'darkOrchid', 'darkRed', 'darkSalmon', 'darkSeaGreen', 'darkSlateBlue',
        'darkSlateGray', 'darkSlateGrey', 'darkTurquoise', 'darkViolet',
        'dkBlue', 'dkCyan', 'dkGoldenrod', 'dkGray', 'dkGrey', 'dkGreen',
        'dkKhaki', 'dkMagenta', 'dkOliveGreen', 'dkOrange', 'dkOrchid', 'dkRed',
        'dkSalmon', 'dkSeaGreen', 'dkSlateBlue', 'dkSlateGray', 'dkSlateGrey',
        'dkTurquoise', 'dkViolet', 'deepPink', 'deepSkyBlue', 'dimGray',
        'dimGrey', 'dodgerBlue', 'firebrick', 'floralWhite', 'forestGreen',
        'fuchsia', 'gainsboro', 'ghostWhite', 'gold', 'goldenrod', 'gray',
        'grey', 'green', 'greenYellow', 'honeydew', 'hotPink', 'indianRed',
        'indigo', 'ivory', 'khaki', 'lavender', 'lavenderBlush', 'lawnGreen',
        'lemonChiffon', 'lightBlue', 'lightCoral', 'lightCyan',
        'lightGoldenrodYellow', 'lightGray', 'lightGrey', 'lightGreen',
        'lightPink', 'lightSalmon', 'lightSeaGreen', 'lightSkyBlue',
        'lightSlateGray', 'lightSlateGrey', 'lightSteelBlue', 'lightYellow',
        'ltBlue', 'ltCoral', 'ltCyan', 'ltGoldenrodYellow', 'ltGray', 'ltGrey',
        'ltGreen', 'ltPink', 'ltSalmon', 'ltSeaGreen', 'ltSkyBlue',
        'ltSlateGray', 'ltSlateGrey', 'ltSteelBlue', 'ltYellow', 'lime',
        'limeGreen', 'linen', 'magenta', 'maroon', 'medAquamarine', 'medBlue',
        'medOrchid', 'medPurple', 'medSeaGreen', 'medSlateBlue',
        'medSpringGreen', 'medTurquoise', 'medVioletRed', 'mediumAquamarine',
        'mediumBlue', 'mediumOrchid', 'mediumPurple', 'mediumSeaGreen',
        'mediumSlateBlue', 'mediumSpringGreen', 'mediumTurquoise',
        'mediumVioletRed', 'midnightBlue', 'mintCream', 'mistyRose', 'moccasin',
        'navajoWhite', 'navy', 'oldLace', 'olive', 'oliveDrab', 'orange',
        'orangeRed', 'orchid', 'paleGoldenrod', 'paleGreen', 'paleTurquoise',
        'paleVioletRed', 'papayaWhip', 'peachPuff', 'peru', 'pink', 'plum',
        'powderBlue', 'purple', 'red', 'rosyBrown', 'royalBlue', 'saddleBrown',
        'salmon', 'sandyBrown', 'seaGreen', 'seaShell', 'sienna', 'silver',
        'skyBlue', 'slateBlue', 'slateGray', 'slateGrey', 'snow', 'springGreen',
        'steelBlue', 'tan', 'teal', 'thistle', 'tomato', 'turquoise', 'violet',
        'wheat', 'white', 'whiteSmoke', 'yellow', 'yellowGreen'
    ]


SCHEME_COLORS= ['bg1', 'tx1', 'bg2', 'tx2', 'accent1', 'accent2', 'accent3',
                'accent4', 'accent5', 'accent6', 'hlink', 'folHlink', 'phClr', 'dk1', 'lt1',
                'dk2', 'lt2'
                ]


class Transform(Serialisable):

    pass


class SystemColor(Serialisable):

    tagname = "sysClr"
    namespace = DRAWING_NS

    # color transform options
    tint = NestedInteger(allow_none=True)
    shade = NestedInteger(allow_none=True)
    comp = Typed(expected_type=Transform, allow_none=True)
    inv = Typed(expected_type=Transform, allow_none=True)
    gray = Typed(expected_type=Transform, allow_none=True)
    alpha = NestedInteger(allow_none=True)
    alphaOff = NestedInteger(allow_none=True)
    alphaMod = NestedInteger(allow_none=True)
    hue = NestedInteger(allow_none=True)
    hueOff = NestedInteger(allow_none=True)
    hueMod = NestedInteger(allow_none=True)
    sat = NestedInteger(allow_none=True)
    satOff = NestedInteger(allow_none=True)
    satMod = NestedInteger(allow_none=True)
    lum = NestedInteger(allow_none=True)
    lumOff = NestedInteger(allow_none=True)
    lumMod = NestedInteger(allow_none=True)
    red = NestedInteger(allow_none=True)
    redOff = NestedInteger(allow_none=True)
    redMod = NestedInteger(allow_none=True)
    green = NestedInteger(allow_none=True)
    greenOff = NestedInteger(allow_none=True)
    greenMod = NestedInteger(allow_none=True)
    blue = NestedInteger(allow_none=True)
    blueOff = NestedInteger(allow_none=True)
    blueMod = NestedInteger(allow_none=True)
    gamma = Typed(expected_type=Transform, allow_none=True)
    invGamma = Typed(expected_type=Transform, allow_none=True)

    val = Set(values=( ['scrollBar', 'background', 'activeCaption',
                        'inactiveCaption', 'menu', 'window', 'windowFrame', 'menuText',
                        'windowText', 'captionText', 'activeBorder', 'inactiveBorder',
                        'appWorkspace', 'highlight', 'highlightText', 'btnFace', 'btnShadow',
                        'grayText', 'btnText', 'inactiveCaptionText', 'btnHighlight',
                        '3dDkShadow', '3dLight', 'infoText', 'infoBk', 'hotLight',
                        'gradientActiveCaption', 'gradientInactiveCaption', 'menuHighlight',
                        'menuBar'] )
              )
    lastClr = RGB(allow_none=True)

    __elements__ = ('tint', 'shade', 'comp', 'inv', 'gray', "alpha",
                    "alphaOff", "alphaMod", "hue", "hueOff", "hueMod", "hueOff", "sat",
                    "satOff", "satMod", "lum", "lumOff", "lumMod", "red", "redOff", "redMod",
                    "green", "greenOff", "greenMod", "blue", "blueOff", "blueMod", "gamma",
                    "invGamma")

    def __init__(self,
                 val="windowText",
                 lastClr=None,
                 tint=None,
                 shade=None,
                 comp=None,
                 inv=None,
                 gray=None,
                 alpha=None,
                 alphaOff=None,
                 alphaMod=None,
                 hue=None,
                 hueOff=None,
                 hueMod=None,
                 sat=None,
                 satOff=None,
                 satMod=None,
                 lum=None,
                 lumOff=None,
                 lumMod=None,
                 red=None,
                 redOff=None,
                 redMod=None,
                 green=None,
                 greenOff=None,
                 greenMod=None,
                 blue=None,
                 blueOff=None,
                 blueMod=None,
                 gamma=None,
                 invGamma=None
                ):
        self.val = val
        self.lastClr = lastClr
        self.tint = tint
        self.shade = shade
        self.comp = comp
        self.inv = inv
        self.gray = gray
        self.alpha = alpha
        self.alphaOff = alphaOff
        self.alphaMod = alphaMod
        self.hue = hue
        self.hueOff = hueOff
        self.hueMod = hueMod
        self.sat = sat
        self.satOff = satOff
        self.satMod = satMod
        self.lum = lum
        self.lumOff = lumOff
        self.lumMod = lumMod
        self.red = red
        self.redOff = redOff
        self.redMod = redMod
        self.green = green
        self.greenOff = greenOff
        self.greenMod = greenMod
        self.blue = blue
        self.blueOff = blueOff
        self.blueMod = blueMod
        self.gamma = gamma
        self.invGamma = invGamma


class HSLColor(Serialisable):

    tagname = "hslClr"

    hue = Integer()
    sat = MinMax(min=0, max=100)
    lum = MinMax(min=0, max=100)

    #TODO add color transform options

    def __init__(self,
                 hue=None,
                 sat=None,
                 lum=None,
                ):
        self.hue = hue
        self.sat = sat
        self.lum = lum



class RGBPercent(Serialisable):

    tagname = "rgbClr"

    r = MinMax(min=0, max=100)
    g = MinMax(min=0, max=100)
    b = MinMax(min=0, max=100)

    #TODO add color transform options

    def __init__(self,
                 r=None,
                 g=None,
                 b=None,
                ):
        self.r = r
        self.g = g
        self.b = b


class SchemeColor(Serialisable):

    tagname = "schemeClr"
    namespace = DRAWING_NS

    tint = NestedInteger(allow_none=True)
    shade = NestedInteger(allow_none=True)
    comp = EmptyTag(allow_none=True)
    inv = NestedInteger(allow_none=True)
    gray = NestedInteger(allow_none=True)
    alpha = NestedInteger(allow_none=True)
    alphaOff = NestedInteger(allow_none=True)
    alphaMod = NestedInteger(allow_none=True)
    hue = NestedInteger(allow_none=True)
    hueOff = NestedInteger(allow_none=True)
    hueMod = NestedInteger(allow_none=True)
    sat = NestedInteger(allow_none=True)
    satOff = NestedInteger(allow_none=True)
    satMod = NestedInteger(allow_none=True)
    lum = NestedInteger(allow_none=True)
    lumOff = NestedInteger(allow_none=True)
    lumMod = NestedInteger(allow_none=True)
    red = NestedInteger(allow_none=True)
    redOff = NestedInteger(allow_none=True)
    redMod = NestedInteger(allow_none=True)
    green = NestedInteger(allow_none=True)
    greenOff = NestedInteger(allow_none=True)
    greenMod = NestedInteger(allow_none=True)
    blue = NestedInteger(allow_none=True)
    blueOff = NestedInteger(allow_none=True)
    blueMod = NestedInteger(allow_none=True)
    gamma = EmptyTag(allow_none=True)
    invGamma = EmptyTag(allow_none=True)
    val = Set(values=(['bg1', 'tx1', 'bg2', 'tx2', 'accent1', 'accent2',
                       'accent3', 'accent4', 'accent5', 'accent6', 'hlink', 'folHlink', 'phClr',
                       'dk1', 'lt1', 'dk2', 'lt2']))

    __elements__ = ('tint', 'shade', 'comp', 'inv', 'gray', 'alpha',
                    'alphaOff', 'alphaMod', 'hue', 'hueOff', 'hueMod', 'sat', 'satOff',
                    'satMod', 'lum', 'lumMod', 'lumOff', 'red', 'redOff', 'redMod', 'green',
                    'greenOff', 'greenMod', 'blue', 'blueOff', 'blueMod', 'gamma',
                    'invGamma')

    def __init__(self,
                 tint=None,
                 shade=None,
                 comp=None,
                 inv=None,
                 gray=None,
                 alpha=None,
                 alphaOff=None,
                 alphaMod=None,
                 hue=None,
                 hueOff=None,
                 hueMod=None,
                 sat=None,
                 satOff=None,
                 satMod=None,
                 lum=None,
                 lumOff=None,
                 lumMod=None,
                 red=None,
                 redOff=None,
                 redMod=None,
                 green=None,
                 greenOff=None,
                 greenMod=None,
                 blue=None,
                 blueOff=None,
                 blueMod=None,
                 gamma=None,
                 invGamma=None,
                 val=None,
                ):
        self.tint = tint
        self.shade = shade
        self.comp = comp
        self.inv = inv
        self.gray = gray
        self.alpha = alpha
        self.alphaOff = alphaOff
        self.alphaMod = alphaMod
        self.hue = hue
        self.hueOff = hueOff
        self.hueMod = hueMod
        self.sat = sat
        self.satOff = satOff
        self.satMod = satMod
        self.lum = lum
        self.lumOff = lumOff
        self.lumMod = lumMod
        self.red = red
        self.redOff = redOff
        self.redMod = redMod
        self.green = green
        self.greenOff = greenOff
        self.greenMod = greenMod
        self.blue = blue
        self.blueOff = blueOff
        self.blueMod = blueMod
        self.gamma = gamma
        self.invGamma = invGamma
        self.val = val

class ColorChoice(Serialisable):

    tagname = "colorChoice"
    namespace = DRAWING_NS

    scrgbClr = Typed(expected_type=RGBPercent, allow_none=True)
    RGBPercent = Alias('scrgbClr')
    srgbClr = NestedValue(expected_type=str, allow_none=True) # needs pattern and can have transform
    RGB = Alias('srgbClr')
    hslClr = Typed(expected_type=HSLColor, allow_none=True)
    sysClr = Typed(expected_type=SystemColor, allow_none=True)
    schemeClr = Typed(expected_type=SchemeColor, allow_none=True)
    prstClr = NestedNoneSet(values=PRESET_COLORS)

    __elements__ = ('scrgbClr', 'srgbClr', 'hslClr', 'sysClr', 'schemeClr', 'prstClr')

    def __init__(self,
                 scrgbClr=None,
                 srgbClr=None,
                 hslClr=None,
                 sysClr=None,
                 schemeClr=None,
                 prstClr=None,
                ):
        self.scrgbClr = scrgbClr
        self.srgbClr = srgbClr
        self.hslClr = hslClr
        self.sysClr = sysClr
        self.schemeClr = schemeClr
        self.prstClr = prstClr

_COLOR_SET = ('dk1', 'lt1', 'dk2', 'lt2', 'accent1', 'accent2', 'accent3',
               'accent4', 'accent5', 'accent6', 'hlink', 'folHlink')


class ColorMapping(Serialisable):

    tagname = "clrMapOvr"

    bg1 = Set(values=_COLOR_SET)
    tx1 = Set(values=_COLOR_SET)
    bg2 = Set(values=_COLOR_SET)
    tx2 = Set(values=_COLOR_SET)
    accent1 = Set(values=_COLOR_SET)
    accent2 = Set(values=_COLOR_SET)
    accent3 = Set(values=_COLOR_SET)
    accent4 = Set(values=_COLOR_SET)
    accent5 = Set(values=_COLOR_SET)
    accent6 = Set(values=_COLOR_SET)
    hlink = Set(values=_COLOR_SET)
    folHlink = Set(values=_COLOR_SET)
    extLst = Typed(expected_type=OfficeArtExtensionList, allow_none=True)

    def __init__(self,
                 bg1="lt1",
                 tx1="dk1",
                 bg2="lt2",
                 tx2="dk2",
                 accent1="accent1",
                 accent2="accent2",
                 accent3="accent3",
                 accent4="accent4",
                 accent5="accent5",
                 accent6="accent6",
                 hlink="hlink",
                 folHlink="folHlink",
                 extLst=None,
                ):
        self.bg1 = bg1
        self.tx1 = tx1
        self.bg2 = bg2
        self.tx2 = tx2
        self.accent1 = accent1
        self.accent2 = accent2
        self.accent3 = accent3
        self.accent4 = accent4
        self.accent5 = accent5
        self.accent6 = accent6
        self.hlink = hlink
        self.folHlink = folHlink
        self.extLst = extLst


class ColorChoiceDescriptor(Typed):
    """
    Objects can choose from 7 different kinds of color system.
    Assume RGBHex if a string is passed in.
    """

    expected_type = ColorChoice
    allow_none = True

    def __set__(self, instance, value):
        if isinstance(value, str):
            value = ColorChoice(srgbClr=value)
        else:
            if hasattr(self, "namespace") and value is not None:
                value.namespace = self.namespace
        super().__set__(instance, value)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\urllib3\util\url.py ===
from __future__ import absolute_import

import re
from collections import namedtuple

from ..exceptions import LocationParseError
from ..packages import six

url_attrs = ["scheme", "auth", "host", "port", "path", "query", "fragment"]

# We only want to normalize urls with an HTTP(S) scheme.
# urllib3 infers URLs without a scheme (None) to be http.
NORMALIZABLE_SCHEMES = ("http", "https", None)

# Almost all of these patterns were derived from the
# 'rfc3986' module: https://github.com/python-hyper/rfc3986
PERCENT_RE = re.compile(r"%[a-fA-F0-9]{2}")
SCHEME_RE = re.compile(r"^(?:[a-zA-Z][a-zA-Z0-9+-]*:|/)")
URI_RE = re.compile(
    r"^(?:([a-zA-Z][a-zA-Z0-9+.-]*):)?"
    r"(?://([^\\/?#]*))?"
    r"([^?#]*)"
    r"(?:\?([^#]*))?"
    r"(?:#(.*))?$",
    re.UNICODE | re.DOTALL,
)

IPV4_PAT = r"(?:[0-9]{1,3}\.){3}[0-9]{1,3}"
HEX_PAT = "[0-9A-Fa-f]{1,4}"
LS32_PAT = "(?:{hex}:{hex}|{ipv4})".format(hex=HEX_PAT, ipv4=IPV4_PAT)
_subs = {"hex": HEX_PAT, "ls32": LS32_PAT}
_variations = [
    #                            6( h16 ":" ) ls32
    "(?:%(hex)s:){6}%(ls32)s",
    #                       "::" 5( h16 ":" ) ls32
    "::(?:%(hex)s:){5}%(ls32)s",
    # [               h16 ] "::" 4( h16 ":" ) ls32
    "(?:%(hex)s)?::(?:%(hex)s:){4}%(ls32)s",
    # [ *1( h16 ":" ) h16 ] "::" 3( h16 ":" ) ls32
    "(?:(?:%(hex)s:)?%(hex)s)?::(?:%(hex)s:){3}%(ls32)s",
    # [ *2( h16 ":" ) h16 ] "::" 2( h16 ":" ) ls32
    "(?:(?:%(hex)s:){0,2}%(hex)s)?::(?:%(hex)s:){2}%(ls32)s",
    # [ *3( h16 ":" ) h16 ] "::"    h16 ":"   ls32
    "(?:(?:%(hex)s:){0,3}%(hex)s)?::%(hex)s:%(ls32)s",
    # [ *4( h16 ":" ) h16 ] "::"              ls32
    "(?:(?:%(hex)s:){0,4}%(hex)s)?::%(ls32)s",
    # [ *5( h16 ":" ) h16 ] "::"              h16
    "(?:(?:%(hex)s:){0,5}%(hex)s)?::%(hex)s",
    # [ *6( h16 ":" ) h16 ] "::"
    "(?:(?:%(hex)s:){0,6}%(hex)s)?::",
]

UNRESERVED_PAT = r"ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789._\-~"
IPV6_PAT = "(?:" + "|".join([x % _subs for x in _variations]) + ")"
ZONE_ID_PAT = "(?:%25|%)(?:[" + UNRESERVED_PAT + "]|%[a-fA-F0-9]{2})+"
IPV6_ADDRZ_PAT = r"\[" + IPV6_PAT + r"(?:" + ZONE_ID_PAT + r")?\]"
REG_NAME_PAT = r"(?:[^\[\]%:/?#]|%[a-fA-F0-9]{2})*"
TARGET_RE = re.compile(r"^(/[^?#]*)(?:\?([^#]*))?(?:#.*)?$")

IPV4_RE = re.compile("^" + IPV4_PAT + "$")
IPV6_RE = re.compile("^" + IPV6_PAT + "$")
IPV6_ADDRZ_RE = re.compile("^" + IPV6_ADDRZ_PAT + "$")
BRACELESS_IPV6_ADDRZ_RE = re.compile("^" + IPV6_ADDRZ_PAT[2:-2] + "$")
ZONE_ID_RE = re.compile("(" + ZONE_ID_PAT + r")\]$")

_HOST_PORT_PAT = ("^(%s|%s|%s)(?::0*?(|0|[1-9][0-9]{0,4}))?$") % (
    REG_NAME_PAT,
    IPV4_PAT,
    IPV6_ADDRZ_PAT,
)
_HOST_PORT_RE = re.compile(_HOST_PORT_PAT, re.UNICODE | re.DOTALL)

UNRESERVED_CHARS = set(
    "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789._-~"
)
SUB_DELIM_CHARS = set("!$&'()*+,;=")
USERINFO_CHARS = UNRESERVED_CHARS | SUB_DELIM_CHARS | {":"}
PATH_CHARS = USERINFO_CHARS | {"@", "/"}
QUERY_CHARS = FRAGMENT_CHARS = PATH_CHARS | {"?"}


class Url(namedtuple("Url", url_attrs)):
    """
    Data structure for representing an HTTP URL. Used as a return value for
    :func:`parse_url`. Both the scheme and host are normalized as they are
    both case-insensitive according to RFC 3986.
    """

    __slots__ = ()

    def __new__(
        cls,
        scheme=None,
        auth=None,
        host=None,
        port=None,
        path=None,
        query=None,
        fragment=None,
    ):
        if path and not path.startswith("/"):
            path = "/" + path
        if scheme is not None:
            scheme = scheme.lower()
        return super(Url, cls).__new__(
            cls, scheme, auth, host, port, path, query, fragment
        )

    @property
    def hostname(self):
        """For backwards-compatibility with urlparse. We're nice like that."""
        return self.host

    @property
    def request_uri(self):
        """Absolute path including the query string."""
        uri = self.path or "/"

        if self.query is not None:
            uri += "?" + self.query

        return uri

    @property
    def netloc(self):
        """Network location including host and port"""
        if self.port:
            return "%s:%d" % (self.host, self.port)
        return self.host

    @property
    def url(self):
        """
        Convert self into a url

        This function should more or less round-trip with :func:`.parse_url`. The
        returned url may not be exactly the same as the url inputted to
        :func:`.parse_url`, but it should be equivalent by the RFC (e.g., urls
        with a blank port will have : removed).

        Example: ::

            >>> U = parse_url('http://google.com/mail/')
            >>> U.url
            'http://google.com/mail/'
            >>> Url('http', 'username:password', 'host.com', 80,
            ... '/path', 'query', 'fragment').url
            'http://username:password@host.com:80/path?query#fragment'
        """
        scheme, auth, host, port, path, query, fragment = self
        url = u""

        # We use "is not None" we want things to happen with empty strings (or 0 port)
        if scheme is not None:
            url += scheme + u"://"
        if auth is not None:
            url += auth + u"@"
        if host is not None:
            url += host
        if port is not None:
            url += u":" + str(port)
        if path is not None:
            url += path
        if query is not None:
            url += u"?" + query
        if fragment is not None:
            url += u"#" + fragment

        return url

    def __str__(self):
        return self.url


def split_first(s, delims):
    """
    .. deprecated:: 1.25

    Given a string and an iterable of delimiters, split on the first found
    delimiter. Return two split parts and the matched delimiter.

    If not found, then the first part is the full input string.

    Example::

        >>> split_first('foo/bar?baz', '?/=')
        ('foo', 'bar?baz', '/')
        >>> split_first('foo/bar?baz', '123')
        ('foo/bar?baz', '', None)

    Scales linearly with number of delims. Not ideal for large number of delims.
    """
    min_idx = None
    min_delim = None
    for d in delims:
        idx = s.find(d)
        if idx < 0:
            continue

        if min_idx is None or idx < min_idx:
            min_idx = idx
            min_delim = d

    if min_idx is None or min_idx < 0:
        return s, "", None

    return s[:min_idx], s[min_idx + 1 :], min_delim


def _encode_invalid_chars(component, allowed_chars, encoding="utf-8"):
    """Percent-encodes a URI component without reapplying
    onto an already percent-encoded component.
    """
    if component is None:
        return component

    component = six.ensure_text(component)

    # Normalize existing percent-encoded bytes.
    # Try to see if the component we're encoding is already percent-encoded
    # so we can skip all '%' characters but still encode all others.
    component, percent_encodings = PERCENT_RE.subn(
        lambda match: match.group(0).upper(), component
    )

    uri_bytes = component.encode("utf-8", "surrogatepass")
    is_percent_encoded = percent_encodings == uri_bytes.count(b"%")
    encoded_component = bytearray()

    for i in range(0, len(uri_bytes)):
        # Will return a single character bytestring on both Python 2 & 3
        byte = uri_bytes[i : i + 1]
        byte_ord = ord(byte)
        if (is_percent_encoded and byte == b"%") or (
            byte_ord < 128 and byte.decode() in allowed_chars
        ):
            encoded_component += byte
            continue
        encoded_component.extend(b"%" + (hex(byte_ord)[2:].encode().zfill(2).upper()))

    return encoded_component.decode(encoding)


def _remove_path_dot_segments(path):
    # See http://tools.ietf.org/html/rfc3986#section-5.2.4 for pseudo-code
    segments = path.split("/")  # Turn the path into a list of segments
    output = []  # Initialize the variable to use to store output

    for segment in segments:
        # '.' is the current directory, so ignore it, it is superfluous
        if segment == ".":
            continue
        # Anything other than '..', should be appended to the output
        elif segment != "..":
            output.append(segment)
        # In this case segment == '..', if we can, we should pop the last
        # element
        elif output:
            output.pop()

    # If the path starts with '/' and the output is empty or the first string
    # is non-empty
    if path.startswith("/") and (not output or output[0]):
        output.insert(0, "")

    # If the path starts with '/.' or '/..' ensure we add one more empty
    # string to add a trailing '/'
    if path.endswith(("/.", "/..")):
        output.append("")

    return "/".join(output)


def _normalize_host(host, scheme):
    if host:
        if isinstance(host, six.binary_type):
            host = six.ensure_str(host)

        if scheme in NORMALIZABLE_SCHEMES:
            is_ipv6 = IPV6_ADDRZ_RE.match(host)
            if is_ipv6:
                # IPv6 hosts of the form 'a::b%zone' are encoded in a URL as
                # such per RFC 6874: 'a::b%25zone'. Unquote the ZoneID
                # separator as necessary to return a valid RFC 4007 scoped IP.
                match = ZONE_ID_RE.search(host)
                if match:
                    start, end = match.span(1)
                    zone_id = host[start:end]

                    if zone_id.startswith("%25") and zone_id != "%25":
                        zone_id = zone_id[3:]
                    else:
                        zone_id = zone_id[1:]
                    zone_id = "%" + _encode_invalid_chars(zone_id, UNRESERVED_CHARS)
                    return host[:start].lower() + zone_id + host[end:]
                else:
                    return host.lower()
            elif not IPV4_RE.match(host):
                return six.ensure_str(
                    b".".join([_idna_encode(label) for label in host.split(".")])
                )
    return host


def _idna_encode(name):
    if name and any(ord(x) >= 128 for x in name):
        try:
            from pip._vendor import idna
        except ImportError:
            six.raise_from(
                LocationParseError("Unable to parse URL without the 'idna' module"),
                None,
            )
        try:
            return idna.encode(name.lower(), strict=True, std3_rules=True)
        except idna.IDNAError:
            six.raise_from(
                LocationParseError(u"Name '%s' is not a valid IDNA label" % name), None
            )
    return name.lower().encode("ascii")


def _encode_target(target):
    """Percent-encodes a request target so that there are no invalid characters"""
    path, query = TARGET_RE.match(target).groups()
    target = _encode_invalid_chars(path, PATH_CHARS)
    query = _encode_invalid_chars(query, QUERY_CHARS)
    if query is not None:
        target += "?" + query
    return target


def parse_url(url):
    """
    Given a url, return a parsed :class:`.Url` namedtuple. Best-effort is
    performed to parse incomplete urls. Fields not provided will be None.
    This parser is RFC 3986 and RFC 6874 compliant.

    The parser logic and helper functions are based heavily on
    work done in the ``rfc3986`` module.

    :param str url: URL to parse into a :class:`.Url` namedtuple.

    Partly backwards-compatible with :mod:`urlparse`.

    Example::

        >>> parse_url('http://google.com/mail/')
        Url(scheme='http', host='google.com', port=None, path='/mail/', ...)
        >>> parse_url('google.com:80')
        Url(scheme=None, host='google.com', port=80, path=None, ...)
        >>> parse_url('/foo?bar')
        Url(scheme=None, host=None, port=None, path='/foo', query='bar', ...)
    """
    if not url:
        # Empty
        return Url()

    source_url = url
    if not SCHEME_RE.search(url):
        url = "//" + url

    try:
        scheme, authority, path, query, fragment = URI_RE.match(url).groups()
        normalize_uri = scheme is None or scheme.lower() in NORMALIZABLE_SCHEMES

        if scheme:
            scheme = scheme.lower()

        if authority:
            auth, _, host_port = authority.rpartition("@")
            auth = auth or None
            host, port = _HOST_PORT_RE.match(host_port).groups()
            if auth and normalize_uri:
                auth = _encode_invalid_chars(auth, USERINFO_CHARS)
            if port == "":
                port = None
        else:
            auth, host, port = None, None, None

        if port is not None:
            port = int(port)
            if not (0 <= port <= 65535):
                raise LocationParseError(url)

        host = _normalize_host(host, scheme)

        if normalize_uri and path:
            path = _remove_path_dot_segments(path)
            path = _encode_invalid_chars(path, PATH_CHARS)
        if normalize_uri and query:
            query = _encode_invalid_chars(query, QUERY_CHARS)
        if normalize_uri and fragment:
            fragment = _encode_invalid_chars(fragment, FRAGMENT_CHARS)

    except (ValueError, AttributeError):
        return six.raise_from(LocationParseError(source_url), None)

    # For the sake of backwards compatibility we put empty
    # string values for path if there are any defined values
    # beyond the path in the URL.
    # TODO: Remove this when we break backwards compatibility.
    if not path:
        if query is not None or fragment is not None:
            path = ""
        else:
            path = None

    # Ensure that each part of the URL is a `str` for
    # backwards compatibility.
    if isinstance(url, six.text_type):
        ensure_func = six.ensure_text
    else:
        ensure_func = six.ensure_str

    def ensure_type(x):
        return x if x is None else ensure_func(x)

    return Url(
        scheme=ensure_type(scheme),
        auth=ensure_type(auth),
        host=ensure_type(host),
        port=port,
        path=ensure_type(path),
        query=ensure_type(query),
        fragment=ensure_type(fragment),
    )


def get_host(url):
    """
    Deprecated. Use :func:`parse_url` instead.
    """
    p = parse_url(url)
    return p.scheme or "http", p.hostname, p.port

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\testing\exclusions.py ===
# testing/exclusions.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors

import contextlib
import operator
import re
import sys

from . import config
from .. import util
from ..util import decorator
from ..util.compat import inspect_getfullargspec


def skip_if(predicate, reason=None):
    rule = compound()
    pred = _as_predicate(predicate, reason)
    rule.skips.add(pred)
    return rule


def fails_if(predicate, reason=None):
    rule = compound()
    pred = _as_predicate(predicate, reason)
    rule.fails.add(pred)
    return rule


class compound:
    def __init__(self):
        self.fails = set()
        self.skips = set()

    def __add__(self, other):
        return self.add(other)

    def as_skips(self):
        rule = compound()
        rule.skips.update(self.skips)
        rule.skips.update(self.fails)
        return rule

    def add(self, *others):
        copy = compound()
        copy.fails.update(self.fails)
        copy.skips.update(self.skips)

        for other in others:
            copy.fails.update(other.fails)
            copy.skips.update(other.skips)
        return copy

    def not_(self):
        copy = compound()
        copy.fails.update(NotPredicate(fail) for fail in self.fails)
        copy.skips.update(NotPredicate(skip) for skip in self.skips)
        return copy

    @property
    def enabled(self):
        return self.enabled_for_config(config._current)

    def enabled_for_config(self, config):
        for predicate in self.skips.union(self.fails):
            if predicate(config):
                return False
        else:
            return True

    def matching_config_reasons(self, config):
        return [
            predicate._as_string(config)
            for predicate in self.skips.union(self.fails)
            if predicate(config)
        ]

    def _extend(self, other):
        self.skips.update(other.skips)
        self.fails.update(other.fails)

    def __call__(self, fn):
        if hasattr(fn, "_sa_exclusion_extend"):
            fn._sa_exclusion_extend._extend(self)
            return fn

        @decorator
        def decorate(fn, *args, **kw):
            return self._do(config._current, fn, *args, **kw)

        decorated = decorate(fn)
        decorated._sa_exclusion_extend = self
        return decorated

    @contextlib.contextmanager
    def fail_if(self):
        all_fails = compound()
        all_fails.fails.update(self.skips.union(self.fails))

        try:
            yield
        except Exception as ex:
            all_fails._expect_failure(config._current, ex)
        else:
            all_fails._expect_success(config._current)

    def _do(self, cfg, fn, *args, **kw):
        for skip in self.skips:
            if skip(cfg):
                msg = "'%s' : %s" % (
                    config.get_current_test_name(),
                    skip._as_string(cfg),
                )
                config.skip_test(msg)

        try:
            return_value = fn(*args, **kw)
        except Exception as ex:
            self._expect_failure(cfg, ex, name=fn.__name__)
        else:
            self._expect_success(cfg, name=fn.__name__)
            return return_value

    def _expect_failure(self, config, ex, name="block"):
        for fail in self.fails:
            if fail(config):
                print(
                    "%s failed as expected (%s): %s "
                    % (name, fail._as_string(config), ex)
                )
                break
        else:
            raise ex.with_traceback(sys.exc_info()[2])

    def _expect_success(self, config, name="block"):
        if not self.fails:
            return

        for fail in self.fails:
            if fail(config):
                raise AssertionError(
                    "Unexpected success for '%s' (%s)"
                    % (
                        name,
                        " and ".join(
                            fail._as_string(config) for fail in self.fails
                        ),
                    )
                )


def only_if(predicate, reason=None):
    predicate = _as_predicate(predicate)
    return skip_if(NotPredicate(predicate), reason)


def succeeds_if(predicate, reason=None):
    predicate = _as_predicate(predicate)
    return fails_if(NotPredicate(predicate), reason)


class Predicate:
    @classmethod
    def as_predicate(cls, predicate, description=None):
        if isinstance(predicate, compound):
            return cls.as_predicate(predicate.enabled_for_config, description)
        elif isinstance(predicate, Predicate):
            if description and predicate.description is None:
                predicate.description = description
            return predicate
        elif isinstance(predicate, (list, set)):
            return OrPredicate(
                [cls.as_predicate(pred) for pred in predicate], description
            )
        elif isinstance(predicate, tuple):
            return SpecPredicate(*predicate)
        elif isinstance(predicate, str):
            tokens = re.match(
                r"([\+\w]+)\s*(?:(>=|==|!=|<=|<|>)\s*([\d\.]+))?", predicate
            )
            if not tokens:
                raise ValueError(
                    "Couldn't locate DB name in predicate: %r" % predicate
                )
            db = tokens.group(1)
            op = tokens.group(2)
            spec = (
                tuple(int(d) for d in tokens.group(3).split("."))
                if tokens.group(3)
                else None
            )

            return SpecPredicate(db, op, spec, description=description)
        elif callable(predicate):
            return LambdaPredicate(predicate, description)
        else:
            assert False, "unknown predicate type: %s" % predicate

    def _format_description(self, config, negate=False):
        bool_ = self(config)
        if negate:
            bool_ = not negate
        return self.description % {
            "driver": (
                config.db.url.get_driver_name() if config else "<no driver>"
            ),
            "database": (
                config.db.url.get_backend_name() if config else "<no database>"
            ),
            "doesnt_support": "doesn't support" if bool_ else "does support",
            "does_support": "does support" if bool_ else "doesn't support",
        }

    def _as_string(self, config=None, negate=False):
        raise NotImplementedError()


class BooleanPredicate(Predicate):
    def __init__(self, value, description=None):
        self.value = value
        self.description = description or "boolean %s" % value

    def __call__(self, config):
        return self.value

    def _as_string(self, config, negate=False):
        return self._format_description(config, negate=negate)


class SpecPredicate(Predicate):
    def __init__(self, db, op=None, spec=None, description=None):
        self.db = db
        self.op = op
        self.spec = spec
        self.description = description

    _ops = {
        "<": operator.lt,
        ">": operator.gt,
        "==": operator.eq,
        "!=": operator.ne,
        "<=": operator.le,
        ">=": operator.ge,
        "in": operator.contains,
        "between": lambda val, pair: val >= pair[0] and val <= pair[1],
    }

    def __call__(self, config):
        if config is None:
            return False

        engine = config.db

        if "+" in self.db:
            dialect, driver = self.db.split("+")
        else:
            dialect, driver = self.db, None

        if dialect and engine.name != dialect:
            return False
        if driver is not None and engine.driver != driver:
            return False

        if self.op is not None:
            assert driver is None, "DBAPI version specs not supported yet"

            version = _server_version(engine)
            oper = (
                hasattr(self.op, "__call__") and self.op or self._ops[self.op]
            )
            return oper(version, self.spec)
        else:
            return True

    def _as_string(self, config, negate=False):
        if self.description is not None:
            return self._format_description(config)
        elif self.op is None:
            if negate:
                return "not %s" % self.db
            else:
                return "%s" % self.db
        else:
            if negate:
                return "not %s %s %s" % (self.db, self.op, self.spec)
            else:
                return "%s %s %s" % (self.db, self.op, self.spec)


class LambdaPredicate(Predicate):
    def __init__(self, lambda_, description=None, args=None, kw=None):
        spec = inspect_getfullargspec(lambda_)
        if not spec[0]:
            self.lambda_ = lambda db: lambda_()
        else:
            self.lambda_ = lambda_
        self.args = args or ()
        self.kw = kw or {}
        if description:
            self.description = description
        elif lambda_.__doc__:
            self.description = lambda_.__doc__
        else:
            self.description = "custom function"

    def __call__(self, config):
        return self.lambda_(config)

    def _as_string(self, config, negate=False):
        return self._format_description(config)


class NotPredicate(Predicate):
    def __init__(self, predicate, description=None):
        self.predicate = predicate
        self.description = description

    def __call__(self, config):
        return not self.predicate(config)

    def _as_string(self, config, negate=False):
        if self.description:
            return self._format_description(config, not negate)
        else:
            return self.predicate._as_string(config, not negate)


class OrPredicate(Predicate):
    def __init__(self, predicates, description=None):
        self.predicates = predicates
        self.description = description

    def __call__(self, config):
        for pred in self.predicates:
            if pred(config):
                return True
        return False

    def _eval_str(self, config, negate=False):
        if negate:
            conjunction = " and "
        else:
            conjunction = " or "
        return conjunction.join(
            p._as_string(config, negate=negate) for p in self.predicates
        )

    def _negation_str(self, config):
        if self.description is not None:
            return "Not " + self._format_description(config)
        else:
            return self._eval_str(config, negate=True)

    def _as_string(self, config, negate=False):
        if negate:
            return self._negation_str(config)
        else:
            if self.description is not None:
                return self._format_description(config)
            else:
                return self._eval_str(config)


_as_predicate = Predicate.as_predicate


def _is_excluded(db, op, spec):
    return SpecPredicate(db, op, spec)(config._current)


def _server_version(engine):
    """Return a server_version_info tuple."""

    # force metadata to be retrieved
    conn = engine.connect()
    version = getattr(engine.dialect, "server_version_info", None)
    if version is None:
        version = ()
    conn.close()
    return version


def db_spec(*dbs):
    return OrPredicate([Predicate.as_predicate(db) for db in dbs])


def open():  # noqa
    return skip_if(BooleanPredicate(False, "mark as execute"))


def closed():
    return skip_if(BooleanPredicate(True, "marked as skip"))


def fails(reason=None):
    return fails_if(BooleanPredicate(True, reason or "expected to fail"))


def future():
    return fails_if(BooleanPredicate(True, "Future feature"))


def fails_on(db, reason=None):
    return fails_if(db, reason)


def fails_on_everything_except(*dbs):
    return succeeds_if(OrPredicate([Predicate.as_predicate(db) for db in dbs]))


def skip(db, reason=None):
    return skip_if(db, reason)


def only_on(dbs, reason=None):
    return only_if(
        OrPredicate(
            [Predicate.as_predicate(db, reason) for db in util.to_list(dbs)]
        )
    )


def exclude(db, op, spec, reason=None):
    return skip_if(SpecPredicate(db, op, spec), reason)


def against(config, *queries):
    assert queries, "no queries sent!"
    return OrPredicate([Predicate.as_predicate(query) for query in queries])(
        config
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\combinatorics\prufer.py ===
from sympy.core import Basic
from sympy.core.containers import Tuple
from sympy.tensor.array import Array
from sympy.core.sympify import _sympify
from sympy.utilities.iterables import flatten, iterable
from sympy.utilities.misc import as_int

from collections import defaultdict


class Prufer(Basic):
    """
    The Prufer correspondence is an algorithm that describes the
    bijection between labeled trees and the Prufer code. A Prufer
    code of a labeled tree is unique up to isomorphism and has
    a length of n - 2.

    Prufer sequences were first used by Heinz Prufer to give a
    proof of Cayley's formula.

    References
    ==========

    .. [1] https://mathworld.wolfram.com/LabeledTree.html

    """
    _prufer_repr = None
    _tree_repr = None
    _nodes = None
    _rank = None

    @property
    def prufer_repr(self):
        """Returns Prufer sequence for the Prufer object.

        This sequence is found by removing the highest numbered vertex,
        recording the node it was attached to, and continuing until only
        two vertices remain. The Prufer sequence is the list of recorded nodes.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> Prufer([[0, 3], [1, 3], [2, 3], [3, 4], [4, 5]]).prufer_repr
        [3, 3, 3, 4]
        >>> Prufer([1, 0, 0]).prufer_repr
        [1, 0, 0]

        See Also
        ========

        to_prufer

        """
        if self._prufer_repr is None:
            self._prufer_repr = self.to_prufer(self._tree_repr[:], self.nodes)
        return self._prufer_repr

    @property
    def tree_repr(self):
        """Returns the tree representation of the Prufer object.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> Prufer([[0, 3], [1, 3], [2, 3], [3, 4], [4, 5]]).tree_repr
        [[0, 3], [1, 3], [2, 3], [3, 4], [4, 5]]
        >>> Prufer([1, 0, 0]).tree_repr
        [[1, 2], [0, 1], [0, 3], [0, 4]]

        See Also
        ========

        to_tree

        """
        if self._tree_repr is None:
            self._tree_repr = self.to_tree(self._prufer_repr[:])
        return self._tree_repr

    @property
    def nodes(self):
        """Returns the number of nodes in the tree.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> Prufer([[0, 3], [1, 3], [2, 3], [3, 4], [4, 5]]).nodes
        6
        >>> Prufer([1, 0, 0]).nodes
        5

        """
        return self._nodes

    @property
    def rank(self):
        """Returns the rank of the Prufer sequence.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> p = Prufer([[0, 3], [1, 3], [2, 3], [3, 4], [4, 5]])
        >>> p.rank
        778
        >>> p.next(1).rank
        779
        >>> p.prev().rank
        777

        See Also
        ========

        prufer_rank, next, prev, size

        """
        if self._rank is None:
            self._rank = self.prufer_rank()
        return self._rank

    @property
    def size(self):
        """Return the number of possible trees of this Prufer object.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> Prufer([0]*4).size == Prufer([6]*4).size == 1296
        True

        See Also
        ========

        prufer_rank, rank, next, prev

        """
        return self.prev(self.rank).prev().rank + 1

    @staticmethod
    def to_prufer(tree, n):
        """Return the Prufer sequence for a tree given as a list of edges where
        ``n`` is the number of nodes in the tree.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> a = Prufer([[0, 1], [0, 2], [0, 3]])
        >>> a.prufer_repr
        [0, 0]
        >>> Prufer.to_prufer([[0, 1], [0, 2], [0, 3]], 4)
        [0, 0]

        See Also
        ========
        prufer_repr: returns Prufer sequence of a Prufer object.

        """
        d = defaultdict(int)
        L = []
        for edge in tree:
            # Increment the value of the corresponding
            # node in the degree list as we encounter an
            # edge involving it.
            d[edge[0]] += 1
            d[edge[1]] += 1
        for i in range(n - 2):
            # find the smallest leaf
            for x in range(n):
                if d[x] == 1:
                    break
            # find the node it was connected to
            y = None
            for edge in tree:
                if x == edge[0]:
                    y = edge[1]
                elif x == edge[1]:
                    y = edge[0]
                if y is not None:
                    break
            # record and update
            L.append(y)
            for j in (x, y):
                d[j] -= 1
                if not d[j]:
                    d.pop(j)
            tree.remove(edge)
        return L

    @staticmethod
    def to_tree(prufer):
        """Return the tree (as a list of edges) of the given Prufer sequence.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> a = Prufer([0, 2], 4)
        >>> a.tree_repr
        [[0, 1], [0, 2], [2, 3]]
        >>> Prufer.to_tree([0, 2])
        [[0, 1], [0, 2], [2, 3]]

        References
        ==========

        .. [1] https://hamberg.no/erlend/posts/2010-11-06-prufer-sequence-compact-tree-representation.html

        See Also
        ========
        tree_repr: returns tree representation of a Prufer object.

        """
        tree = []
        last = []
        n = len(prufer) + 2
        d = defaultdict(lambda: 1)
        for p in prufer:
            d[p] += 1
        for i in prufer:
            for j in range(n):
            # find the smallest leaf (degree = 1)
                if d[j] == 1:
                    break
            # (i, j) is the new edge that we append to the tree
            # and remove from the degree dictionary
            d[i] -= 1
            d[j] -= 1
            tree.append(sorted([i, j]))
        last = [i for i in range(n) if d[i] == 1] or [0, 1]
        tree.append(last)

        return tree

    @staticmethod
    def edges(*runs):
        """Return a list of edges and the number of nodes from the given runs
        that connect nodes in an integer-labelled tree.

        All node numbers will be shifted so that the minimum node is 0. It is
        not a problem if edges are repeated in the runs; only unique edges are
        returned. There is no assumption made about what the range of the node
        labels should be, but all nodes from the smallest through the largest
        must be present.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> Prufer.edges([1, 2, 3], [2, 4, 5]) # a T
        ([[0, 1], [1, 2], [1, 3], [3, 4]], 5)

        Duplicate edges are removed:

        >>> Prufer.edges([0, 1, 2, 3], [1, 4, 5], [1, 4, 6]) # a K
        ([[0, 1], [1, 2], [1, 4], [2, 3], [4, 5], [4, 6]], 7)

        """
        e = set()
        nmin = runs[0][0]
        for r in runs:
            for i in range(len(r) - 1):
                a, b = r[i: i + 2]
                if b < a:
                    a, b = b, a
                e.add((a, b))
        rv = []
        got = set()
        nmin = nmax = None
        for ei in e:
            got.update(ei)
            nmin = min(ei[0], nmin) if nmin is not None else ei[0]
            nmax = max(ei[1], nmax) if nmax is not None else ei[1]
            rv.append(list(ei))
        missing = set(range(nmin, nmax + 1)) - got
        if missing:
            missing = [i + nmin for i in missing]
            if len(missing) == 1:
                msg = 'Node %s is missing.' % missing.pop()
            else:
                msg = 'Nodes %s are missing.' % sorted(missing)
            raise ValueError(msg)
        if nmin != 0:
            for i, ei in enumerate(rv):
                rv[i] = [n - nmin for n in ei]
            nmax -= nmin
        return sorted(rv), nmax + 1

    def prufer_rank(self):
        """Computes the rank of a Prufer sequence.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> a = Prufer([[0, 1], [0, 2], [0, 3]])
        >>> a.prufer_rank()
        0

        See Also
        ========

        rank, next, prev, size

        """
        r = 0
        p = 1
        for i in range(self.nodes - 3, -1, -1):
            r += p*self.prufer_repr[i]
            p *= self.nodes
        return r

    @classmethod
    def unrank(self, rank, n):
        """Finds the unranked Prufer sequence.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> Prufer.unrank(0, 4)
        Prufer([0, 0])

        """
        n, rank = as_int(n), as_int(rank)
        L = defaultdict(int)
        for i in range(n - 3, -1, -1):
            L[i] = rank % n
            rank = (rank - L[i])//n
        return Prufer([L[i] for i in range(len(L))])

    def __new__(cls, *args, **kw_args):
        """The constructor for the Prufer object.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer

        A Prufer object can be constructed from a list of edges:

        >>> a = Prufer([[0, 1], [0, 2], [0, 3]])
        >>> a.prufer_repr
        [0, 0]

        If the number of nodes is given, no checking of the nodes will
        be performed; it will be assumed that nodes 0 through n - 1 are
        present:

        >>> Prufer([[0, 1], [0, 2], [0, 3]], 4)
        Prufer([[0, 1], [0, 2], [0, 3]], 4)

        A Prufer object can be constructed from a Prufer sequence:

        >>> b = Prufer([1, 3])
        >>> b.tree_repr
        [[0, 1], [1, 3], [2, 3]]

        """
        arg0 = Array(args[0]) if args[0] else Tuple()
        args = (arg0,) + tuple(_sympify(arg) for arg in args[1:])
        ret_obj = Basic.__new__(cls, *args, **kw_args)
        args = [list(args[0])]
        if args[0] and iterable(args[0][0]):
            if not args[0][0]:
                raise ValueError(
                    'Prufer expects at least one edge in the tree.')
            if len(args) > 1:
                nnodes = args[1]
            else:
                nodes = set(flatten(args[0]))
                nnodes = max(nodes) + 1
                if nnodes != len(nodes):
                    missing = set(range(nnodes)) - nodes
                    if len(missing) == 1:
                        msg = 'Node %s is missing.' % missing.pop()
                    else:
                        msg = 'Nodes %s are missing.' % sorted(missing)
                    raise ValueError(msg)
            ret_obj._tree_repr = [list(i) for i in args[0]]
            ret_obj._nodes = nnodes
        else:
            ret_obj._prufer_repr = args[0]
            ret_obj._nodes = len(ret_obj._prufer_repr) + 2
        return ret_obj

    def next(self, delta=1):
        """Generates the Prufer sequence that is delta beyond the current one.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> a = Prufer([[0, 1], [0, 2], [0, 3]])
        >>> b = a.next(1) # == a.next()
        >>> b.tree_repr
        [[0, 2], [0, 1], [1, 3]]
        >>> b.rank
        1

        See Also
        ========

        prufer_rank, rank, prev, size

        """
        return Prufer.unrank(self.rank + delta, self.nodes)

    def prev(self, delta=1):
        """Generates the Prufer sequence that is -delta before the current one.

        Examples
        ========

        >>> from sympy.combinatorics.prufer import Prufer
        >>> a = Prufer([[0, 1], [1, 2], [2, 3], [1, 4]])
        >>> a.rank
        36
        >>> b = a.prev()
        >>> b
        Prufer([1, 2, 0])
        >>> b.rank
        35

        See Also
        ========

        prufer_rank, rank, next, size

        """
        return Prufer.unrank(self.rank -delta, self.nodes)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\solvers.py ===
"""Low-level linear systems solver. """


from sympy.utilities.exceptions import sympy_deprecation_warning
from sympy.utilities.iterables import connected_components

from sympy.core.sympify import sympify
from sympy.core.numbers import Integer, Rational
from sympy.matrices.dense import MutableDenseMatrix
from sympy.polys.domains import ZZ, QQ

from sympy.polys.domains import EX
from sympy.polys.rings import sring
from sympy.polys.polyerrors import NotInvertible
from sympy.polys.domainmatrix import DomainMatrix


class PolyNonlinearError(Exception):
    """Raised by solve_lin_sys for nonlinear equations"""
    pass


class RawMatrix(MutableDenseMatrix):
    """
    .. deprecated:: 1.9

       This class fundamentally is broken by design. Use ``DomainMatrix`` if
       you want a matrix over the polys domains or ``Matrix`` for a matrix
       with ``Expr`` elements. The ``RawMatrix`` class will be removed/broken
       in future in order to reestablish the invariant that the elements of a
       Matrix should be of type ``Expr``.

    """
    _sympify = staticmethod(lambda x, *args, **kwargs: x)

    def __init__(self, *args, **kwargs):
        sympy_deprecation_warning(
            """
            The RawMatrix class is deprecated. Use either DomainMatrix or
            Matrix instead.
            """,
            deprecated_since_version="1.9",
            active_deprecations_target="deprecated-rawmatrix",
        )

        domain = ZZ
        for i in range(self.rows):
            for j in range(self.cols):
                val = self[i,j]
                if getattr(val, 'is_Poly', False):
                    K = val.domain[val.gens]
                    val_sympy = val.as_expr()
                elif hasattr(val, 'parent'):
                    K = val.parent()
                    val_sympy = K.to_sympy(val)
                elif isinstance(val, (int, Integer)):
                    K = ZZ
                    val_sympy = sympify(val)
                elif isinstance(val, Rational):
                    K = QQ
                    val_sympy = val
                else:
                    for K in ZZ, QQ:
                        if K.of_type(val):
                            val_sympy = K.to_sympy(val)
                            break
                    else:
                        raise TypeError
                domain = domain.unify(K)
                self[i,j] = val_sympy
        self.ring = domain


def eqs_to_matrix(eqs_coeffs, eqs_rhs, gens, domain):
    """Get matrix from linear equations in dict format.

    Explanation
    ===========

    Get the matrix representation of a system of linear equations represented
    as dicts with low-level DomainElement coefficients. This is an
    *internal* function that is used by solve_lin_sys.

    Parameters
    ==========

    eqs_coeffs: list[dict[Symbol, DomainElement]]
        The left hand sides of the equations as dicts mapping from symbols to
        coefficients where the coefficients are instances of
        DomainElement.
    eqs_rhs: list[DomainElements]
        The right hand sides of the equations as instances of
        DomainElement.
    gens: list[Symbol]
        The unknowns in the system of equations.
    domain: Domain
        The domain for coefficients of both lhs and rhs.

    Returns
    =======

    The augmented matrix representation of the system as a DomainMatrix.

    Examples
    ========

    >>> from sympy import symbols, ZZ
    >>> from sympy.polys.solvers import eqs_to_matrix
    >>> x, y = symbols('x, y')
    >>> eqs_coeff = [{x:ZZ(1), y:ZZ(1)}, {x:ZZ(1), y:ZZ(-1)}]
    >>> eqs_rhs = [ZZ(0), ZZ(-1)]
    >>> eqs_to_matrix(eqs_coeff, eqs_rhs, [x, y], ZZ)
    DomainMatrix([[1, 1, 0], [1, -1, 1]], (2, 3), ZZ)

    See also
    ========

    solve_lin_sys: Uses :func:`~eqs_to_matrix` internally
    """
    sym2index = {x: n for n, x in enumerate(gens)}
    nrows = len(eqs_coeffs)
    ncols = len(gens) + 1
    rows = [[domain.zero] * ncols for _ in range(nrows)]
    for row, eq_coeff, eq_rhs in zip(rows, eqs_coeffs, eqs_rhs):
        for sym, coeff in eq_coeff.items():
            row[sym2index[sym]] = domain.convert(coeff)
        row[-1] = -domain.convert(eq_rhs)

    return DomainMatrix(rows, (nrows, ncols), domain)


def sympy_eqs_to_ring(eqs, symbols):
    """Convert a system of equations from Expr to a PolyRing

    Explanation
    ===========

    High-level functions like ``solve`` expect Expr as inputs but can use
    ``solve_lin_sys`` internally. This function converts equations from
    ``Expr`` to the low-level poly types used by the ``solve_lin_sys``
    function.

    Parameters
    ==========

    eqs: List of Expr
        A list of equations as Expr instances
    symbols: List of Symbol
        A list of the symbols that are the unknowns in the system of
        equations.

    Returns
    =======

    Tuple[List[PolyElement], Ring]: The equations as PolyElement instances
    and the ring of polynomials within which each equation is represented.

    Examples
    ========

    >>> from sympy import symbols
    >>> from sympy.polys.solvers import sympy_eqs_to_ring
    >>> a, x, y = symbols('a, x, y')
    >>> eqs = [x-y, x+a*y]
    >>> eqs_ring, ring = sympy_eqs_to_ring(eqs, [x, y])
    >>> eqs_ring
    [x - y, x + a*y]
    >>> type(eqs_ring[0])
    <class 'sympy.polys.rings.PolyElement'>
    >>> ring
    ZZ(a)[x,y]

    With the equations in this form they can be passed to ``solve_lin_sys``:

    >>> from sympy.polys.solvers import solve_lin_sys
    >>> solve_lin_sys(eqs_ring, ring)
    {y: 0, x: 0}
    """
    try:
        K, eqs_K = sring(eqs, symbols, field=True, extension=True)
    except NotInvertible:
        # https://github.com/sympy/sympy/issues/18874
        K, eqs_K = sring(eqs, symbols, domain=EX)
    return eqs_K, K.to_domain()


def solve_lin_sys(eqs, ring, _raw=True):
    """Solve a system of linear equations from a PolynomialRing

    Explanation
    ===========

    Solves a system of linear equations given as PolyElement instances of a
    PolynomialRing. The basic arithmetic is carried out using instance of
    DomainElement which is more efficient than :class:`~sympy.core.expr.Expr`
    for the most common inputs.

    While this is a public function it is intended primarily for internal use
    so its interface is not necessarily convenient. Users are suggested to use
    the :func:`sympy.solvers.solveset.linsolve` function (which uses this
    function internally) instead.

    Parameters
    ==========

    eqs: list[PolyElement]
        The linear equations to be solved as elements of a
        PolynomialRing (assumed equal to zero).
    ring: PolynomialRing
        The polynomial ring from which eqs are drawn. The generators of this
        ring are the unknowns to be solved for and the domain of the ring is
        the domain of the coefficients of the system of equations.
    _raw: bool
        If *_raw* is False, the keys and values in the returned dictionary
        will be of type Expr (and the unit of the field will be removed from
        the keys) otherwise the low-level polys types will be returned, e.g.
        PolyElement: PythonRational.

    Returns
    =======

    ``None`` if the system has no solution.

    dict[Symbol, Expr] if _raw=False

    dict[Symbol, DomainElement] if _raw=True.

    Examples
    ========

    >>> from sympy import symbols
    >>> from sympy.polys.solvers import solve_lin_sys, sympy_eqs_to_ring
    >>> x, y = symbols('x, y')
    >>> eqs = [x - y, x + y - 2]
    >>> eqs_ring, ring = sympy_eqs_to_ring(eqs, [x, y])
    >>> solve_lin_sys(eqs_ring, ring)
    {y: 1, x: 1}

    Passing ``_raw=False`` returns the same result except that the keys are
    ``Expr`` rather than low-level poly types.

    >>> solve_lin_sys(eqs_ring, ring, _raw=False)
    {x: 1, y: 1}

    See also
    ========

    sympy_eqs_to_ring: prepares the inputs to ``solve_lin_sys``.
    linsolve: ``linsolve`` uses ``solve_lin_sys`` internally.
    sympy.solvers.solvers.solve: ``solve`` uses ``solve_lin_sys`` internally.
    """
    as_expr = not _raw

    assert ring.domain.is_Field

    eqs_dict = [dict(eq) for eq in eqs]

    one_monom = ring.one.monoms()[0]
    zero = ring.domain.zero

    eqs_rhs = []
    eqs_coeffs = []
    for eq_dict in eqs_dict:
        eq_rhs = eq_dict.pop(one_monom, zero)
        eq_coeffs = {}
        for monom, coeff in eq_dict.items():
            if sum(monom) != 1:
                msg = "Nonlinear term encountered in solve_lin_sys"
                raise PolyNonlinearError(msg)
            eq_coeffs[ring.gens[monom.index(1)]] = coeff
        if not eq_coeffs:
            if not eq_rhs:
                continue
            else:
                return None
        eqs_rhs.append(eq_rhs)
        eqs_coeffs.append(eq_coeffs)

    result = _solve_lin_sys(eqs_coeffs, eqs_rhs, ring)

    if result is not None and as_expr:

        def to_sympy(x):
            as_expr = getattr(x, 'as_expr', None)
            if as_expr:
                return as_expr()
            else:
                return ring.domain.to_sympy(x)

        tresult = {to_sympy(sym): to_sympy(val) for sym, val in result.items()}

        # Remove 1.0x
        result = {}
        for k, v in tresult.items():
            if k.is_Mul:
                c, s = k.as_coeff_Mul()
                result[s] = v/c
            else:
                result[k] = v

    return result


def _solve_lin_sys(eqs_coeffs, eqs_rhs, ring):
    """Solve a linear system from dict of PolynomialRing coefficients

    Explanation
    ===========

    This is an **internal** function used by :func:`solve_lin_sys` after the
    equations have been preprocessed. The role of this function is to split
    the system into connected components and pass those to
    :func:`_solve_lin_sys_component`.

    Examples
    ========

    Setup a system for $x-y=0$ and $x+y=2$ and solve:

    >>> from sympy import symbols, sring
    >>> from sympy.polys.solvers import _solve_lin_sys
    >>> x, y = symbols('x, y')
    >>> R, (xr, yr) = sring([x, y], [x, y])
    >>> eqs = [{xr:R.one, yr:-R.one}, {xr:R.one, yr:R.one}]
    >>> eqs_rhs = [R.zero, -2*R.one]
    >>> _solve_lin_sys(eqs, eqs_rhs, R)
    {y: 1, x: 1}

    See also
    ========

    solve_lin_sys: This function is used internally by :func:`solve_lin_sys`.
    """
    V = ring.gens
    E = []
    for eq_coeffs in eqs_coeffs:
        syms = list(eq_coeffs)
        E.extend(zip(syms[:-1], syms[1:]))
    G = V, E

    components = connected_components(G)

    sym2comp = {}
    for n, component in enumerate(components):
        for sym in component:
            sym2comp[sym] = n

    subsystems = [([], []) for _ in range(len(components))]
    for eq_coeff, eq_rhs in zip(eqs_coeffs, eqs_rhs):
        sym = next(iter(eq_coeff), None)
        sub_coeff, sub_rhs = subsystems[sym2comp[sym]]
        sub_coeff.append(eq_coeff)
        sub_rhs.append(eq_rhs)

    sol = {}
    for subsystem in subsystems:
        subsol = _solve_lin_sys_component(subsystem[0], subsystem[1], ring)
        if subsol is None:
            return None
        sol.update(subsol)

    return sol


def _solve_lin_sys_component(eqs_coeffs, eqs_rhs, ring):
    """Solve a linear system from dict of PolynomialRing coefficients

    Explanation
    ===========

    This is an **internal** function used by :func:`solve_lin_sys` after the
    equations have been preprocessed. After :func:`_solve_lin_sys` splits the
    system into connected components this function is called for each
    component. The system of equations is solved using Gauss-Jordan
    elimination with division followed by back-substitution.

    Examples
    ========

    Setup a system for $x-y=0$ and $x+y=2$ and solve:

    >>> from sympy import symbols, sring
    >>> from sympy.polys.solvers import _solve_lin_sys_component
    >>> x, y = symbols('x, y')
    >>> R, (xr, yr) = sring([x, y], [x, y])
    >>> eqs = [{xr:R.one, yr:-R.one}, {xr:R.one, yr:R.one}]
    >>> eqs_rhs = [R.zero, -2*R.one]
    >>> _solve_lin_sys_component(eqs, eqs_rhs, R)
    {y: 1, x: 1}

    See also
    ========

    solve_lin_sys: This function is used internally by :func:`solve_lin_sys`.
    """

    # transform from equations to matrix form
    matrix = eqs_to_matrix(eqs_coeffs, eqs_rhs, ring.gens, ring.domain)

    # convert to a field for rref
    if not matrix.domain.is_Field:
        matrix = matrix.to_field()

    # solve by row-reduction
    echelon, pivots = matrix.rref()

    # construct the returnable form of the solutions
    keys = ring.gens

    if pivots and pivots[-1] == len(keys):
        return None

    if len(pivots) == len(keys):
        sol = []
        for s in [row[-1] for row in echelon.rep.to_ddm()]:
            a = s
            sol.append(a)
        sols = dict(zip(keys, sol))
    else:
        sols = {}
        g = ring.gens
        # Extract ground domain coefficients and convert to the ring:
        if hasattr(ring, 'ring'):
            convert = ring.ring.ground_new
        else:
            convert = ring.ground_new
        echelon = echelon.rep.to_ddm()
        vals_set = {v for row in echelon for v in row}
        vals_map = {v: convert(v) for v in vals_set}
        echelon = [[vals_map[eij] for eij in ei] for ei in echelon]
        for i, p in enumerate(pivots):
            v = echelon[i][-1] - sum(echelon[i][j]*g[j] for j in range(p+1, len(g)) if echelon[i][j])
            sols[keys[p]] = v

    return sols

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\profiler.py ===
import argparse
import os

import numpy
import psutil
from onnx import TensorProto

"""
This profiler tool could run a transformer model and print out the kernel time spent on each Node of the model.
Example of profiling of longformer model:
    python profiler.py --model longformer-base-4096_fp32.onnx --batch_size 1 --sequence_length 4096 --global_length 8 --samples 1000 --thread_num 8 --dummy_inputs longformer --use_gpu
Example of importing profile result file from onnxruntime_perf_test:
    python profiler.py --input profile_2021-10-25_12-02-41.json
"""


def parse_arguments(argv=None):
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "-i",
        "--input",
        required=False,
        type=str,
        help="Set the input file for reading the profile results",
    )

    parser.add_argument(
        "-m",
        "--model",
        required=False,
        type=str,
        help="onnx model path to run profiling. Required when --input is not specified.",
    )

    parser.add_argument(
        "-b",
        "--batch_size",
        required=False,
        type=int,
        default=1,
        help="batch size of input",
    )

    parser.add_argument(
        "-s",
        "--sequence_length",
        required=False,
        type=int,
        default=32,
        help="sequence length of input",
    )

    parser.add_argument(
        "--past_sequence_length",
        required=False,
        type=int,
        default=1,
        help="past sequence length for gpt2",
    )

    parser.add_argument(
        "--global_length",
        required=False,
        type=int,
        default=1,
        help="number of global tokens for longformer",
    )

    parser.add_argument(
        "--samples",
        required=False,
        type=int,
        default=1000,
        help="number of samples to test. Set it large enough to reduce the variance of performance result.",
    )

    parser.add_argument(
        "--threshold",
        required=False,
        type=float,
        default=0.01,
        help="Threshold of run time ratio among all nodes. Nodes with larger ratio will show in top expensive nodes.",
    )

    parser.add_argument(
        "--thread_num",
        required=False,
        type=int,
        default=-1,
        help="number of threads to use",
    )

    parser.add_argument(
        "--input_ids_name",
        required=False,
        type=str,
        default=None,
        help="input name for input IDs, for bert",
    )
    parser.add_argument(
        "--segment_ids_name",
        required=False,
        type=str,
        default=None,
        help="input name for segment IDs, for bert",
    )
    parser.add_argument(
        "--input_mask_name",
        required=False,
        type=str,
        default=None,
        help="input name for attention mask, for bert",
    )

    parser.add_argument(
        "--dummy_inputs",
        required=False,
        default="default",
        choices=["bert", "gpt2", "longformer", "default"],
        help="Type of model inputs. The default will create dummy inputs with ones.",
    )

    parser.add_argument("-g", "--use_gpu", required=False, action="store_true", help="use GPU")
    parser.set_defaults(use_gpu=False)

    parser.add_argument(
        "--provider",
        required=False,
        type=str,
        default="cuda",
        help="Execution provider to use",
    )

    parser.add_argument(
        "--basic_optimization",
        required=False,
        action="store_true",
        help="Enable only basic graph optimizations. By default, all optimizations are enabled in OnnxRuntime",
    )
    parser.set_defaults(basic_optimization=False)

    parser.add_argument(
        "--kernel_time_only",
        required=False,
        action="store_true",
        help="Only include the kernel time and no fence time",
    )
    parser.set_defaults(kernel_time_only=False)

    parser.add_argument("-v", "--verbose", required=False, action="store_true")
    parser.set_defaults(verbose=False)

    return parser.parse_args(argv)


def run_profile(onnx_model_path, use_gpu, provider, basic_optimization, thread_num, all_inputs):
    from benchmark_helper import create_onnxruntime_session

    session = create_onnxruntime_session(
        onnx_model_path,
        use_gpu,
        provider,
        enable_all_optimization=not basic_optimization,
        num_threads=thread_num,
        enable_profiling=True,
    )

    for inputs in all_inputs:
        _ = session.run(None, inputs)

    profile_file = session.end_profiling()
    return profile_file


def get_dim_from_type_proto(dim):
    return getattr(dim, dim.WhichOneof("value")) if type(dim.WhichOneof("value")) == str else None  # noqa: E721


def get_shape_from_type_proto(type_proto):
    return [get_dim_from_type_proto(d) for d in type_proto.tensor_type.shape.dim]


def create_dummy_inputs(onnx_model, batch_size, sequence_length, samples):
    """Create dummy inputs for ONNX model.

    Args:
        onnx_model (OnnxModel): ONNX model
        batch_size (int): batch size
        sequence_length (int): sequence length
        samples (int): number of samples

    Returns:
        List[Dict]: list of inputs
    """
    dummy_inputs = {}
    for graph_input in onnx_model.get_graph_inputs_excluding_initializers():
        shape = get_shape_from_type_proto(graph_input.type)
        symbol_dims = []
        for i, dim in enumerate(shape):
            if isinstance(dim, str):
                symbol_dims.append(i)

        # allowed symbolic dimensions: batch_size and sequence_length
        if len(symbol_dims) > 2:
            return None
        if len(symbol_dims) > 0:
            shape[symbol_dims[0]] = batch_size
        if len(symbol_dims) > 1:
            shape[symbol_dims[1]] = sequence_length

        elem_type = graph_input.type.tensor_type.elem_type
        assert elem_type in [TensorProto.FLOAT, TensorProto.INT32, TensorProto.INT64]
        data_type = (
            numpy.float32
            if elem_type == TensorProto.FLOAT
            else (numpy.int64 if elem_type == TensorProto.INT64 else numpy.int32)
        )
        data = numpy.ones(shape, dtype=data_type)
        dummy_inputs[graph_input.name] = data

    all_inputs = [dummy_inputs for _ in range(samples)]
    return all_inputs


def create_bert_inputs(
    onnx_model,
    batch_size,
    sequence_length,
    samples,
    input_ids_name=None,
    segment_ids_name=None,
    input_mask_name=None,
):
    """Create dummy inputs for BERT model.

    Args:
        onnx_model (OnnxModel): ONNX model
        batch_size (int): batch size
        sequence_length (int): sequence length
        samples (int): number of samples
        input_ids_name (str, optional): Name of graph input for input IDs. Defaults to None.
        segment_ids_name (str, optional): Name of graph input for segment IDs. Defaults to None.
        input_mask_name (str, optional): Name of graph input for attention mask. Defaults to None.

    Returns:
        List[Dict]: list of inputs
    """
    from bert_test_data import find_bert_inputs, generate_test_data

    input_ids, segment_ids, input_mask = find_bert_inputs(onnx_model, input_ids_name, segment_ids_name, input_mask_name)
    all_inputs = generate_test_data(
        batch_size,
        sequence_length,
        test_cases=samples,
        seed=123,
        verbose=False,
        input_ids=input_ids,
        segment_ids=segment_ids,
        input_mask=input_mask,
        random_mask_length=False,
    )

    return all_inputs


def create_gpt2_inputs(onnx_model, batch_size, sequence_length, past_sequence_length, samples):
    """Create dummy inputs for GPT-2 model.

    Args:
        onnx_model (OnnxModel): ONNX model
        batch_size (int): batch size
        sequence_length (int): sequence length
        past_sequence_length (int): past sequence length
        samples (int): number of samples

    Raises:
        RuntimeError: symbolic is not supported. Use the tool convert_to_onnx.py to export ONNX model instead.

    Returns:
        List[Dict]: list of inputs
    """
    # The symbolic names shall be same as those used in Gpt2Helper.export_onnx(...) function.
    symbols = {
        "batch_size": batch_size,
        "seq_len": sequence_length,
        "past_seq_len": past_sequence_length,
        "total_seq_len": sequence_length + past_sequence_length,
    }

    dummy_inputs = {}
    for graph_input in onnx_model.get_graph_inputs_excluding_initializers():
        shape = get_shape_from_type_proto(graph_input.type)
        for i, dim in enumerate(shape):
            if isinstance(dim, str):
                if dim not in symbols:
                    raise RuntimeError(f"symbol is not supported: {dim}")
                else:
                    shape[i] = symbols[dim]

        elem_type = graph_input.type.tensor_type.elem_type
        assert elem_type in [TensorProto.FLOAT, TensorProto.INT32, TensorProto.INT64]
        data_type = (
            numpy.float32
            if elem_type == TensorProto.FLOAT
            else (numpy.int64 if elem_type == TensorProto.INT64 else numpy.int32)
        )
        data = numpy.ones(shape, dtype=data_type)
        dummy_inputs[graph_input.name] = data

    all_inputs = [dummy_inputs for _ in range(samples)]
    return all_inputs


def create_longformer_inputs(onnx_model, batch_size, sequence_length, global_length, samples):
    """Create dummy inputs for Longformer model.

    Args:
        onnx_model (OnnxModel): ONNX model
        batch_size (int): batch size
        sequence_length (int): sequence length
        global_length (int): number of global tokens
        samples (int): number of samples

    Raises:
        RuntimeError: symbolic is not supported. Use the tool convert_longformer_to_onnx.py to export ONNX model instead.

    Returns:
        List[Dict]: list of inputs
    """
    symbols = {"batch_size": batch_size, "sequence_length": sequence_length}

    dummy_inputs = {}
    for graph_input in onnx_model.get_graph_inputs_excluding_initializers():
        shape = get_shape_from_type_proto(graph_input.type)
        for i, dim in enumerate(shape):
            if isinstance(dim, str):
                if dim not in symbols:
                    raise RuntimeError(f"symbol is not supported: {dim}")
                else:
                    shape[i] = symbols[dim]

        elem_type = graph_input.type.tensor_type.elem_type
        assert elem_type in [TensorProto.FLOAT, TensorProto.INT32, TensorProto.INT64]
        data_type = (
            numpy.float32
            if elem_type == TensorProto.FLOAT
            else (numpy.int64 if elem_type == TensorProto.INT64 else numpy.int32)
        )

        if "global" in graph_input.name:
            data = numpy.zeros(shape, dtype=data_type)
            data[:, :global_length] = 1
        else:
            data = numpy.ones(shape, dtype=data_type)
        dummy_inputs[graph_input.name] = data

    all_inputs = [dummy_inputs for _ in range(samples)]
    return all_inputs


def run(args):
    num_threads = args.thread_num if args.thread_num > 0 else psutil.cpu_count(logical=False)

    # Set OMP environment variable before importing onnxruntime. Needed for cpu only, and no impact for onnxruntime-gpu package.
    if "OMP_NUM_THREADS" not in os.environ:
        os.environ["OMP_NUM_THREADS"] = str(num_threads)

    from onnx import load
    from onnx_model import OnnxModel

    onnx_model = OnnxModel(load(args.model))

    all_inputs = None
    if args.dummy_inputs == "bert":
        all_inputs = create_bert_inputs(
            onnx_model,
            args.batch_size,
            args.sequence_length,
            args.samples,
            args.input_ids_name,
            args.segment_ids_name,
            args.input_mask_name,
        )
    elif args.dummy_inputs == "gpt2":
        all_inputs = create_gpt2_inputs(
            onnx_model,
            args.batch_size,
            args.sequence_length,
            args.past_sequence_length,
            args.samples,
        )
    elif args.dummy_inputs == "longformer":
        all_inputs = create_longformer_inputs(
            onnx_model,
            args.batch_size,
            args.sequence_length,
            args.global_length,
            args.samples,
        )
    else:  # default
        all_inputs = create_dummy_inputs(onnx_model, args.batch_size, args.sequence_length, args.samples)

    profile_file = run_profile(
        args.model,
        args.use_gpu,
        args.provider,
        args.basic_optimization,
        args.thread_num,
        all_inputs,
    )

    return profile_file


if __name__ == "__main__":
    arguments = parse_arguments()
    print("Arguments", arguments)

    from benchmark_helper import setup_logger

    setup_logger(arguments.verbose)

    if not arguments.input:
        assert arguments.model, "requires either --model to run profiling or --input to read profiling results"
        profile_file = run(arguments)
    else:
        profile_file = arguments.input
    from profile_result_processor import process_results

    results = process_results(profile_file, arguments)

    for line in results:
        print(line)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\expressions\kronecker.py ===
"""Implementation of the Kronecker product"""
from functools import reduce
from math import prod

from sympy.core import Mul, sympify
from sympy.functions import adjoint
from sympy.matrices.exceptions import ShapeError
from sympy.matrices.expressions.matexpr import MatrixExpr
from sympy.matrices.expressions.transpose import transpose
from sympy.matrices.expressions.special import Identity
from sympy.matrices.matrixbase import MatrixBase
from sympy.strategies import (
    canon, condition, distribute, do_one, exhaust, flatten, typed, unpack)
from sympy.strategies.traverse import bottom_up
from sympy.utilities import sift

from .matadd import MatAdd
from .matmul import MatMul
from .matpow import MatPow


def kronecker_product(*matrices):
    """
    The Kronecker product of two or more arguments.

    This computes the explicit Kronecker product for subclasses of
    ``MatrixBase`` i.e. explicit matrices. Otherwise, a symbolic
    ``KroneckerProduct`` object is returned.


    Examples
    ========

    For ``MatrixSymbol`` arguments a ``KroneckerProduct`` object is returned.
    Elements of this matrix can be obtained by indexing, or for MatrixSymbols
    with known dimension the explicit matrix can be obtained with
    ``.as_explicit()``

    >>> from sympy import kronecker_product, MatrixSymbol
    >>> A = MatrixSymbol('A', 2, 2)
    >>> B = MatrixSymbol('B', 2, 2)
    >>> kronecker_product(A)
    A
    >>> kronecker_product(A, B)
    KroneckerProduct(A, B)
    >>> kronecker_product(A, B)[0, 1]
    A[0, 0]*B[0, 1]
    >>> kronecker_product(A, B).as_explicit()
    Matrix([
        [A[0, 0]*B[0, 0], A[0, 0]*B[0, 1], A[0, 1]*B[0, 0], A[0, 1]*B[0, 1]],
        [A[0, 0]*B[1, 0], A[0, 0]*B[1, 1], A[0, 1]*B[1, 0], A[0, 1]*B[1, 1]],
        [A[1, 0]*B[0, 0], A[1, 0]*B[0, 1], A[1, 1]*B[0, 0], A[1, 1]*B[0, 1]],
        [A[1, 0]*B[1, 0], A[1, 0]*B[1, 1], A[1, 1]*B[1, 0], A[1, 1]*B[1, 1]]])

    For explicit matrices the Kronecker product is returned as a Matrix

    >>> from sympy import Matrix, kronecker_product
    >>> sigma_x = Matrix([
    ... [0, 1],
    ... [1, 0]])
    ...
    >>> Isigma_y = Matrix([
    ... [0, 1],
    ... [-1, 0]])
    ...
    >>> kronecker_product(sigma_x, Isigma_y)
    Matrix([
    [ 0, 0,  0, 1],
    [ 0, 0, -1, 0],
    [ 0, 1,  0, 0],
    [-1, 0,  0, 0]])

    See Also
    ========
        KroneckerProduct

    """
    if not matrices:
        raise TypeError("Empty Kronecker product is undefined")
    if len(matrices) == 1:
        return matrices[0]
    else:
        return KroneckerProduct(*matrices).doit()


class KroneckerProduct(MatrixExpr):
    """
    The Kronecker product of two or more arguments.

    The Kronecker product is a non-commutative product of matrices.
    Given two matrices of dimension (m, n) and (s, t) it produces a matrix
    of dimension (m s, n t).

    This is a symbolic object that simply stores its argument without
    evaluating it. To actually compute the product, use the function
    ``kronecker_product()`` or call the ``.doit()`` or  ``.as_explicit()``
    methods.

    >>> from sympy import KroneckerProduct, MatrixSymbol
    >>> A = MatrixSymbol('A', 5, 5)
    >>> B = MatrixSymbol('B', 5, 5)
    >>> isinstance(KroneckerProduct(A, B), KroneckerProduct)
    True
    """
    is_KroneckerProduct = True

    def __new__(cls, *args, check=True):
        args = list(map(sympify, args))
        if all(a.is_Identity for a in args):
            ret = Identity(prod(a.rows for a in args))
            if all(isinstance(a, MatrixBase) for a in args):
                return ret.as_explicit()
            else:
                return ret

        if check:
            validate(*args)
        return super().__new__(cls, *args)

    @property
    def shape(self):
        rows, cols = self.args[0].shape
        for mat in self.args[1:]:
            rows *= mat.rows
            cols *= mat.cols
        return (rows, cols)

    def _entry(self, i, j, **kwargs):
        result = 1
        for mat in reversed(self.args):
            i, m = divmod(i, mat.rows)
            j, n = divmod(j, mat.cols)
            result *= mat[m, n]
        return result

    def _eval_adjoint(self):
        return KroneckerProduct(*list(map(adjoint, self.args))).doit()

    def _eval_conjugate(self):
        return KroneckerProduct(*[a.conjugate() for a in self.args]).doit()

    def _eval_transpose(self):
        return KroneckerProduct(*list(map(transpose, self.args))).doit()

    def _eval_trace(self):
        from .trace import trace
        return Mul(*[trace(a) for a in self.args])

    def _eval_determinant(self):
        from .determinant import det, Determinant
        if not all(a.is_square for a in self.args):
            return Determinant(self)

        m = self.rows
        return Mul(*[det(a)**(m/a.rows) for a in self.args])

    def _eval_inverse(self):
        try:
            return KroneckerProduct(*[a.inverse() for a in self.args])
        except ShapeError:
            from sympy.matrices.expressions.inverse import Inverse
            return Inverse(self)

    def structurally_equal(self, other):
        '''Determine whether two matrices have the same Kronecker product structure

        Examples
        ========

        >>> from sympy import KroneckerProduct, MatrixSymbol, symbols
        >>> m, n = symbols(r'm, n', integer=True)
        >>> A = MatrixSymbol('A', m, m)
        >>> B = MatrixSymbol('B', n, n)
        >>> C = MatrixSymbol('C', m, m)
        >>> D = MatrixSymbol('D', n, n)
        >>> KroneckerProduct(A, B).structurally_equal(KroneckerProduct(C, D))
        True
        >>> KroneckerProduct(A, B).structurally_equal(KroneckerProduct(D, C))
        False
        >>> KroneckerProduct(A, B).structurally_equal(C)
        False
        '''
        # Inspired by BlockMatrix
        return (isinstance(other, KroneckerProduct)
                and self.shape == other.shape
                and len(self.args) == len(other.args)
                and all(a.shape == b.shape for (a, b) in zip(self.args, other.args)))

    def has_matching_shape(self, other):
        '''Determine whether two matrices have the appropriate structure to bring matrix
        multiplication inside the KroneckerProdut

        Examples
        ========
        >>> from sympy import KroneckerProduct, MatrixSymbol, symbols
        >>> m, n = symbols(r'm, n', integer=True)
        >>> A = MatrixSymbol('A', m, n)
        >>> B = MatrixSymbol('B', n, m)
        >>> KroneckerProduct(A, B).has_matching_shape(KroneckerProduct(B, A))
        True
        >>> KroneckerProduct(A, B).has_matching_shape(KroneckerProduct(A, B))
        False
        >>> KroneckerProduct(A, B).has_matching_shape(A)
        False
        '''
        return (isinstance(other, KroneckerProduct)
                and self.cols == other.rows
                and len(self.args) == len(other.args)
                and all(a.cols == b.rows for (a, b) in zip(self.args, other.args)))

    def _eval_expand_kroneckerproduct(self, **hints):
        return flatten(canon(typed({KroneckerProduct: distribute(KroneckerProduct, MatAdd)}))(self))

    def _kronecker_add(self, other):
        if self.structurally_equal(other):
            return self.__class__(*[a + b for (a, b) in zip(self.args, other.args)])
        else:
            return self + other

    def _kronecker_mul(self, other):
        if self.has_matching_shape(other):
            return self.__class__(*[a*b for (a, b) in zip(self.args, other.args)])
        else:
            return self * other

    def doit(self, **hints):
        deep = hints.get('deep', True)
        if deep:
            args = [arg.doit(**hints) for arg in self.args]
        else:
            args = self.args
        return canonicalize(KroneckerProduct(*args))


def validate(*args):
    if not all(arg.is_Matrix for arg in args):
        raise TypeError("Mix of Matrix and Scalar symbols")


# rules

def extract_commutative(kron):
    c_part = []
    nc_part = []
    for arg in kron.args:
        c, nc = arg.args_cnc()
        c_part.extend(c)
        nc_part.append(Mul._from_args(nc))

    c_part = Mul(*c_part)
    if c_part != 1:
        return c_part*KroneckerProduct(*nc_part)
    return kron


def matrix_kronecker_product(*matrices):
    """Compute the Kronecker product of a sequence of SymPy Matrices.

    This is the standard Kronecker product of matrices [1].

    Parameters
    ==========

    matrices : tuple of MatrixBase instances
        The matrices to take the Kronecker product of.

    Returns
    =======

    matrix : MatrixBase
        The Kronecker product matrix.

    Examples
    ========

    >>> from sympy import Matrix
    >>> from sympy.matrices.expressions.kronecker import (
    ... matrix_kronecker_product)

    >>> m1 = Matrix([[1,2],[3,4]])
    >>> m2 = Matrix([[1,0],[0,1]])
    >>> matrix_kronecker_product(m1, m2)
    Matrix([
    [1, 0, 2, 0],
    [0, 1, 0, 2],
    [3, 0, 4, 0],
    [0, 3, 0, 4]])
    >>> matrix_kronecker_product(m2, m1)
    Matrix([
    [1, 2, 0, 0],
    [3, 4, 0, 0],
    [0, 0, 1, 2],
    [0, 0, 3, 4]])

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Kronecker_product
    """
    # Make sure we have a sequence of Matrices
    if not all(isinstance(m, MatrixBase) for m in matrices):
        raise TypeError(
            'Sequence of Matrices expected, got: %s' % repr(matrices)
        )

    # Pull out the first element in the product.
    matrix_expansion = matrices[-1]
    # Do the kronecker product working from right to left.
    for mat in reversed(matrices[:-1]):
        rows = mat.rows
        cols = mat.cols
        # Go through each row appending kronecker product to.
        # running matrix_expansion.
        for i in range(rows):
            start = matrix_expansion*mat[i*cols]
            # Go through each column joining each item
            for j in range(cols - 1):
                start = start.row_join(
                    matrix_expansion*mat[i*cols + j + 1]
                )
            # If this is the first element, make it the start of the
            # new row.
            if i == 0:
                next = start
            else:
                next = next.col_join(start)
        matrix_expansion = next

    MatrixClass = max(matrices, key=lambda M: M._class_priority).__class__
    if isinstance(matrix_expansion, MatrixClass):
        return matrix_expansion
    else:
        return MatrixClass(matrix_expansion)


def explicit_kronecker_product(kron):
    # Make sure we have a sequence of Matrices
    if not all(isinstance(m, MatrixBase) for m in kron.args):
        return kron

    return matrix_kronecker_product(*kron.args)


rules = (unpack,
         explicit_kronecker_product,
         flatten,
         extract_commutative)

canonicalize = exhaust(condition(lambda x: isinstance(x, KroneckerProduct),
                                 do_one(*rules)))


def _kronecker_dims_key(expr):
    if isinstance(expr, KroneckerProduct):
        return tuple(a.shape for a in expr.args)
    else:
        return (0,)


def kronecker_mat_add(expr):
    args = sift(expr.args, _kronecker_dims_key)
    nonkrons = args.pop((0,), None)
    if not args:
        return expr

    krons = [reduce(lambda x, y: x._kronecker_add(y), group)
             for group in args.values()]

    if not nonkrons:
        return MatAdd(*krons)
    else:
        return MatAdd(*krons) + nonkrons


def kronecker_mat_mul(expr):
    # modified from block matrix code
    factor, matrices = expr.as_coeff_matrices()

    i = 0
    while i < len(matrices) - 1:
        A, B = matrices[i:i+2]
        if isinstance(A, KroneckerProduct) and isinstance(B, KroneckerProduct):
            matrices[i] = A._kronecker_mul(B)
            matrices.pop(i+1)
        else:
            i += 1

    return factor*MatMul(*matrices)


def kronecker_mat_pow(expr):
    if isinstance(expr.base, KroneckerProduct) and all(a.is_square for a in expr.base.args):
        return KroneckerProduct(*[MatPow(a, expr.exp) for a in expr.base.args])
    else:
        return expr


def combine_kronecker(expr):
    """Combine KronekeckerProduct with expression.

    If possible write operations on KroneckerProducts of compatible shapes
    as a single KroneckerProduct.

    Examples
    ========

    >>> from sympy.matrices.expressions import combine_kronecker
    >>> from sympy import MatrixSymbol, KroneckerProduct, symbols
    >>> m, n = symbols(r'm, n', integer=True)
    >>> A = MatrixSymbol('A', m, n)
    >>> B = MatrixSymbol('B', n, m)
    >>> combine_kronecker(KroneckerProduct(A, B)*KroneckerProduct(B, A))
    KroneckerProduct(A*B, B*A)
    >>> combine_kronecker(KroneckerProduct(A, B)+KroneckerProduct(B.T, A.T))
    KroneckerProduct(A + B.T, B + A.T)
    >>> C = MatrixSymbol('C', n, n)
    >>> D = MatrixSymbol('D', m, m)
    >>> combine_kronecker(KroneckerProduct(C, D)**m)
    KroneckerProduct(C**m, D**m)
    """
    def haskron(expr):
        return isinstance(expr, MatrixExpr) and expr.has(KroneckerProduct)

    rule = exhaust(
        bottom_up(exhaust(condition(haskron, typed(
            {MatAdd: kronecker_mat_add,
             MatMul: kronecker_mat_mul,
             MatPow: kronecker_mat_pow})))))
    result = rule(expr)
    doit = getattr(result, 'doit', None)
    if doit is not None:
        return doit()
    else:
        return result

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\printing\printer.py ===
"""Printing subsystem driver

SymPy's printing system works the following way: Any expression can be
passed to a designated Printer who then is responsible to return an
adequate representation of that expression.

**The basic concept is the following:**

1.  Let the object print itself if it knows how.
2.  Take the best fitting method defined in the printer.
3.  As fall-back use the emptyPrinter method for the printer.

Which Method is Responsible for Printing?
^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

The whole printing process is started by calling ``.doprint(expr)`` on the printer
which you want to use. This method looks for an appropriate method which can
print the given expression in the given style that the printer defines.
While looking for the method, it follows these steps:

1.  **Let the object print itself if it knows how.**

    The printer looks for a specific method in every object. The name of that method
    depends on the specific printer and is defined under ``Printer.printmethod``.
    For example, StrPrinter calls ``_sympystr`` and LatexPrinter calls ``_latex``.
    Look at the documentation of the printer that you want to use.
    The name of the method is specified there.

    This was the original way of doing printing in sympy. Every class had
    its own latex, mathml, str and repr methods, but it turned out that it
    is hard to produce a high quality printer, if all the methods are spread
    out that far. Therefore all printing code was combined into the different
    printers, which works great for built-in SymPy objects, but not that
    good for user defined classes where it is inconvenient to patch the
    printers.

2.  **Take the best fitting method defined in the printer.**

    The printer loops through expr classes (class + its bases), and tries
    to dispatch the work to ``_print_<EXPR_CLASS>``

    e.g., suppose we have the following class hierarchy::

            Basic
            |
            Atom
            |
            Number
            |
        Rational

    then, for ``expr=Rational(...)``, the Printer will try
    to call printer methods in the order as shown in the figure below::

        p._print(expr)
        |
        |-- p._print_Rational(expr)
        |
        |-- p._print_Number(expr)
        |
        |-- p._print_Atom(expr)
        |
        `-- p._print_Basic(expr)

    if ``._print_Rational`` method exists in the printer, then it is called,
    and the result is returned back. Otherwise, the printer tries to call
    ``._print_Number`` and so on.

3.  **As a fall-back use the emptyPrinter method for the printer.**

    As fall-back ``self.emptyPrinter`` will be called with the expression. If
    not defined in the Printer subclass this will be the same as ``str(expr)``.

.. _printer_example:

Example of Custom Printer
^^^^^^^^^^^^^^^^^^^^^^^^^

In the example below, we have a printer which prints the derivative of a function
in a shorter form.

.. code-block:: python

    from sympy.core.symbol import Symbol
    from sympy.printing.latex import LatexPrinter, print_latex
    from sympy.core.function import UndefinedFunction, Function


    class MyLatexPrinter(LatexPrinter):
        \"\"\"Print derivative of a function of symbols in a shorter form.
        \"\"\"
        def _print_Derivative(self, expr):
            function, *vars = expr.args
            if not isinstance(type(function), UndefinedFunction) or \\
               not all(isinstance(i, Symbol) for i in vars):
                return super()._print_Derivative(expr)

            # If you want the printer to work correctly for nested
            # expressions then use self._print() instead of str() or latex().
            # See the example of nested modulo below in the custom printing
            # method section.
            return "{}_{{{}}}".format(
                self._print(Symbol(function.func.__name__)),
                            ''.join(self._print(i) for i in vars))


    def print_my_latex(expr):
        \"\"\" Most of the printers define their own wrappers for print().
        These wrappers usually take printer settings. Our printer does not have
        any settings.
        \"\"\"
        print(MyLatexPrinter().doprint(expr))


    y = Symbol("y")
    x = Symbol("x")
    f = Function("f")
    expr = f(x, y).diff(x, y)

    # Print the expression using the normal latex printer and our custom
    # printer.
    print_latex(expr)
    print_my_latex(expr)

The output of the code above is::

    \\frac{\\partial^{2}}{\\partial x\\partial y}  f{\\left(x,y \\right)}
    f_{xy}

.. _printer_method_example:

Example of Custom Printing Method
^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

In the example below, the latex printing of the modulo operator is modified.
This is done by overriding the method ``_latex`` of ``Mod``.

>>> from sympy import Symbol, Mod, Integer, print_latex

>>> # Always use printer._print()
>>> class ModOp(Mod):
...     def _latex(self, printer):
...         a, b = [printer._print(i) for i in self.args]
...         return r"\\operatorname{Mod}{\\left(%s, %s\\right)}" % (a, b)

Comparing the output of our custom operator to the builtin one:

>>> x = Symbol('x')
>>> m = Symbol('m')
>>> print_latex(Mod(x, m))
x \\bmod m
>>> print_latex(ModOp(x, m))
\\operatorname{Mod}{\\left(x, m\\right)}

Common mistakes
~~~~~~~~~~~~~~~
It's important to always use ``self._print(obj)`` to print subcomponents of
an expression when customizing a printer. Mistakes include:

1.  Using ``self.doprint(obj)`` instead:

    >>> # This example does not work properly, as only the outermost call may use
    >>> # doprint.
    >>> class ModOpModeWrong(Mod):
    ...     def _latex(self, printer):
    ...         a, b = [printer.doprint(i) for i in self.args]
    ...         return r"\\operatorname{Mod}{\\left(%s, %s\\right)}" % (a, b)

    This fails when the ``mode`` argument is passed to the printer:

    >>> print_latex(ModOp(x, m), mode='inline')  # ok
    $\\operatorname{Mod}{\\left(x, m\\right)}$
    >>> print_latex(ModOpModeWrong(x, m), mode='inline')  # bad
    $\\operatorname{Mod}{\\left($x$, $m$\\right)}$

2.  Using ``str(obj)`` instead:

    >>> class ModOpNestedWrong(Mod):
    ...     def _latex(self, printer):
    ...         a, b = [str(i) for i in self.args]
    ...         return r"\\operatorname{Mod}{\\left(%s, %s\\right)}" % (a, b)

    This fails on nested objects:

    >>> # Nested modulo.
    >>> print_latex(ModOp(ModOp(x, m), Integer(7)))  # ok
    \\operatorname{Mod}{\\left(\\operatorname{Mod}{\\left(x, m\\right)}, 7\\right)}
    >>> print_latex(ModOpNestedWrong(ModOpNestedWrong(x, m), Integer(7)))  # bad
    \\operatorname{Mod}{\\left(ModOpNestedWrong(x, m), 7\\right)}

3.  Using ``LatexPrinter()._print(obj)`` instead.

    >>> from sympy.printing.latex import LatexPrinter
    >>> class ModOpSettingsWrong(Mod):
    ...     def _latex(self, printer):
    ...         a, b = [LatexPrinter()._print(i) for i in self.args]
    ...         return r"\\operatorname{Mod}{\\left(%s, %s\\right)}" % (a, b)

    This causes all the settings to be discarded in the subobjects. As an
    example, the ``full_prec`` setting which shows floats to full precision is
    ignored:

    >>> from sympy import Float
    >>> print_latex(ModOp(Float(1) * x, m), full_prec=True)  # ok
    \\operatorname{Mod}{\\left(1.00000000000000 x, m\\right)}
    >>> print_latex(ModOpSettingsWrong(Float(1) * x, m), full_prec=True)  # bad
    \\operatorname{Mod}{\\left(1.0 x, m\\right)}

"""

from __future__ import annotations
import sys
from typing import Any, Type
import inspect
from contextlib import contextmanager
from functools import cmp_to_key, update_wrapper

from sympy.core.add import Add
from sympy.core.basic import Basic

from sympy.core.function import AppliedUndef, UndefinedFunction, Function



@contextmanager
def printer_context(printer, **kwargs):
    original = printer._context.copy()
    try:
        printer._context.update(kwargs)
        yield
    finally:
        printer._context = original


class Printer:
    """ Generic printer

    Its job is to provide infrastructure for implementing new printers easily.

    If you want to define your custom Printer or your custom printing method
    for your custom class then see the example above: printer_example_ .
    """

    _global_settings: dict[str, Any] = {}

    _default_settings: dict[str, Any] = {}

    # must be initialized to pass tests and cannot be set to '| None' to pass mypy
    printmethod = None  # type: str

    @classmethod
    def _get_initial_settings(cls):
        settings = cls._default_settings.copy()
        for key, val in cls._global_settings.items():
            if key in cls._default_settings:
                settings[key] = val
        return settings

    def __init__(self, settings=None):
        self._str = str

        self._settings = self._get_initial_settings()
        self._context = {}  # mutable during printing

        if settings is not None:
            self._settings.update(settings)

            if len(self._settings) > len(self._default_settings):
                for key in self._settings:
                    if key not in self._default_settings:
                        raise TypeError("Unknown setting '%s'." % key)

        # _print_level is the number of times self._print() was recursively
        # called. See StrPrinter._print_Float() for an example of usage
        self._print_level = 0

    @classmethod
    def set_global_settings(cls, **settings):
        """Set system-wide printing settings. """
        for key, val in settings.items():
            if val is not None:
                cls._global_settings[key] = val

    @property
    def order(self):
        if 'order' in self._settings:
            return self._settings['order']
        else:
            raise AttributeError("No order defined.")

    def doprint(self, expr):
        """Returns printer's representation for expr (as a string)"""
        return self._str(self._print(expr))

    def _print(self, expr, **kwargs) -> str:
        """Internal dispatcher

        Tries the following concepts to print an expression:
            1. Let the object print itself if it knows how.
            2. Take the best fitting method defined in the printer.
            3. As fall-back use the emptyPrinter method for the printer.
        """
        self._print_level += 1
        try:
            # If the printer defines a name for a printing method
            # (Printer.printmethod) and the object knows for itself how it
            # should be printed, use that method.
            if self.printmethod and hasattr(expr, self.printmethod):
                if not (isinstance(expr, type) and issubclass(expr, Basic)):
                    return getattr(expr, self.printmethod)(self, **kwargs)

            # See if the class of expr is known, or if one of its super
            # classes is known, and use that print function
            # Exception: ignore the subclasses of Undefined, so that, e.g.,
            # Function('gamma') does not get dispatched to _print_gamma
            classes = type(expr).__mro__
            if AppliedUndef in classes:
                classes = classes[classes.index(AppliedUndef):]
            if UndefinedFunction in classes:
                classes = classes[classes.index(UndefinedFunction):]
            # Another exception: if someone subclasses a known function, e.g.,
            # gamma, and changes the name, then ignore _print_gamma
            if Function in classes:
                i = classes.index(Function)
                classes = tuple(c for c in classes[:i] if \
                    c.__name__ == classes[0].__name__ or \
                    c.__name__.endswith("Base")) + classes[i:]
            for cls in classes:
                printmethodname = '_print_' + cls.__name__
                printmethod = getattr(self, printmethodname, None)
                if printmethod is not None:
                    return printmethod(expr, **kwargs)
            # Unknown object, fall back to the emptyPrinter.
            return self.emptyPrinter(expr)
        finally:
            self._print_level -= 1

    def emptyPrinter(self, expr):
        return str(expr)

    def _as_ordered_terms(self, expr, order=None):
        """A compatibility function for ordering terms in Add. """
        order = order or self.order

        if order == 'old':
            return sorted(Add.make_args(expr), key=cmp_to_key(self._compare_pretty))
        elif order == 'none':
            return list(expr.args)
        else:
            return expr.as_ordered_terms(order=order)

    def _compare_pretty(self, a, b):
        """return -1, 0, 1 if a is canonically less, equal or
        greater than b. This is used when 'order=old' is selected
        for printing. This puts Order last, orders Rationals
        according to value, puts terms in order wrt the power of
        the last power appearing in a term. Ties are broken using
        Basic.compare.
        """
        from sympy.core.numbers import Rational
        from sympy.core.symbol import Wild
        from sympy.series.order import Order
        if isinstance(a, Order) and not isinstance(b, Order):
            return 1
        if not isinstance(a, Order) and isinstance(b, Order):
            return -1

        if isinstance(a, Rational) and isinstance(b, Rational):
            l = a.p * b.q
            r = b.p * a.q
            return (l > r) - (l < r)
        else:
            p1, p2, p3 = Wild("p1"), Wild("p2"), Wild("p3")
            r_a = a.match(p1 * p2**p3)
            if r_a and p3 in r_a:
                a3 = r_a[p3]
                r_b = b.match(p1 * p2**p3)
                if r_b and p3 in r_b:
                    b3 = r_b[p3]
                    c = Basic.compare(a3, b3)
                    if c != 0:
                        return c

        # break ties
        return Basic.compare(a, b)


class _PrintFunction:
    """
    Function wrapper to replace ``**settings`` in the signature with printer defaults
    """
    def __init__(self, f, print_cls: Type[Printer]):
        # find all the non-setting arguments
        params = list(inspect.signature(f).parameters.values())
        assert params.pop(-1).kind == inspect.Parameter.VAR_KEYWORD
        self.__other_params = params

        self.__print_cls = print_cls
        update_wrapper(self, f)

    def __reduce__(self):
        # Since this is used as a decorator, it replaces the original function.
        # The default pickling will try to pickle self.__wrapped__ and fail
        # because the wrapped function can't be retrieved by name.
        return self.__wrapped__.__qualname__

    def __call__(self, *args, **kwargs):
        return self.__wrapped__(*args, **kwargs)

    @property
    def __signature__(self) -> inspect.Signature:
        settings = self.__print_cls._get_initial_settings()
        return inspect.Signature(
            parameters=self.__other_params + [
                inspect.Parameter(k, inspect.Parameter.KEYWORD_ONLY, default=v)
                for k, v in settings.items()
            ],
            return_annotation=self.__wrapped__.__annotations__.get('return', inspect.Signature.empty)  # type:ignore
        )


def print_function(print_cls):
    """ A decorator to replace kwargs with the printer settings in __signature__ """
    def decorator(f):
        if sys.version_info < (3, 9):
            # We have to create a subclass so that `help` actually shows the docstring in older Python versions.
            # IPython and Sphinx do not need this, only a raw Python console.
            cls = type(f'{f.__qualname__}_PrintFunction', (_PrintFunction,), {"__doc__": f.__doc__})
        else:
            cls = _PrintFunction
        return cls(f, print_cls)
    return decorator

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\combinatorics\graycode.py ===
from sympy.core import Basic, Integer

import random


class GrayCode(Basic):
    """
    A Gray code is essentially a Hamiltonian walk on
    a n-dimensional cube with edge length of one.
    The vertices of the cube are represented by vectors
    whose values are binary. The Hamilton walk visits
    each vertex exactly once. The Gray code for a 3d
    cube is ['000','100','110','010','011','111','101',
    '001'].

    A Gray code solves the problem of sequentially
    generating all possible subsets of n objects in such
    a way that each subset is obtained from the previous
    one by either deleting or adding a single object.
    In the above example, 1 indicates that the object is
    present, and 0 indicates that its absent.

    Gray codes have applications in statistics as well when
    we want to compute various statistics related to subsets
    in an efficient manner.

    Examples
    ========

    >>> from sympy.combinatorics import GrayCode
    >>> a = GrayCode(3)
    >>> list(a.generate_gray())
    ['000', '001', '011', '010', '110', '111', '101', '100']
    >>> a = GrayCode(4)
    >>> list(a.generate_gray())
    ['0000', '0001', '0011', '0010', '0110', '0111', '0101', '0100', \
    '1100', '1101', '1111', '1110', '1010', '1011', '1001', '1000']

    References
    ==========

    .. [1] Nijenhuis,A. and Wilf,H.S.(1978).
           Combinatorial Algorithms. Academic Press.
    .. [2] Knuth, D. (2011). The Art of Computer Programming, Vol 4
           Addison Wesley


    """

    _skip = False
    _current = 0
    _rank = None

    def __new__(cls, n, *args, **kw_args):
        """
        Default constructor.

        It takes a single argument ``n`` which gives the dimension of the Gray
        code. The starting Gray code string (``start``) or the starting ``rank``
        may also be given; the default is to start at rank = 0 ('0...0').

        Examples
        ========

        >>> from sympy.combinatorics import GrayCode
        >>> a = GrayCode(3)
        >>> a
        GrayCode(3)
        >>> a.n
        3

        >>> a = GrayCode(3, start='100')
        >>> a.current
        '100'

        >>> a = GrayCode(4, rank=4)
        >>> a.current
        '0110'
        >>> a.rank
        4

        """
        if n < 1 or int(n) != n:
            raise ValueError(
                'Gray code dimension must be a positive integer, not %i' % n)
        n = Integer(n)
        args = (n,) + args
        obj = Basic.__new__(cls, *args)
        if 'start' in kw_args:
            obj._current = kw_args["start"]
            if len(obj._current) > n:
                raise ValueError('Gray code start has length %i but '
                'should not be greater than %i' % (len(obj._current), n))
        elif 'rank' in kw_args:
            if int(kw_args["rank"]) != kw_args["rank"]:
                raise ValueError('Gray code rank must be a positive integer, '
                'not %i' % kw_args["rank"])
            obj._rank = int(kw_args["rank"]) % obj.selections
            obj._current = obj.unrank(n, obj._rank)
        return obj

    def next(self, delta=1):
        """
        Returns the Gray code a distance ``delta`` (default = 1) from the
        current value in canonical order.


        Examples
        ========

        >>> from sympy.combinatorics import GrayCode
        >>> a = GrayCode(3, start='110')
        >>> a.next().current
        '111'
        >>> a.next(-1).current
        '010'
        """
        return GrayCode(self.n, rank=(self.rank + delta) % self.selections)

    @property
    def selections(self):
        """
        Returns the number of bit vectors in the Gray code.

        Examples
        ========

        >>> from sympy.combinatorics import GrayCode
        >>> a = GrayCode(3)
        >>> a.selections
        8
        """
        return 2**self.n

    @property
    def n(self):
        """
        Returns the dimension of the Gray code.

        Examples
        ========

        >>> from sympy.combinatorics import GrayCode
        >>> a = GrayCode(5)
        >>> a.n
        5
        """
        return self.args[0]

    def generate_gray(self, **hints):
        """
        Generates the sequence of bit vectors of a Gray Code.

        Examples
        ========

        >>> from sympy.combinatorics import GrayCode
        >>> a = GrayCode(3)
        >>> list(a.generate_gray())
        ['000', '001', '011', '010', '110', '111', '101', '100']
        >>> list(a.generate_gray(start='011'))
        ['011', '010', '110', '111', '101', '100']
        >>> list(a.generate_gray(rank=4))
        ['110', '111', '101', '100']

        See Also
        ========

        skip

        References
        ==========

        .. [1] Knuth, D. (2011). The Art of Computer Programming,
               Vol 4, Addison Wesley

        """
        bits = self.n
        start = None
        if "start" in hints:
            start = hints["start"]
        elif "rank" in hints:
            start = GrayCode.unrank(self.n, hints["rank"])
        if start is not None:
            self._current = start
        current = self.current
        graycode_bin = gray_to_bin(current)
        if len(graycode_bin) > self.n:
            raise ValueError('Gray code start has length %i but should '
            'not be greater than %i' % (len(graycode_bin), bits))
        self._current = int(current, 2)
        graycode_int = int(''.join(graycode_bin), 2)
        for i in range(graycode_int, 1 << bits):
            if self._skip:
                self._skip = False
            else:
                yield self.current
            bbtc = (i ^ (i + 1))
            gbtc = (bbtc ^ (bbtc >> 1))
            self._current = (self._current ^ gbtc)
        self._current = 0

    def skip(self):
        """
        Skips the bit generation.

        Examples
        ========

        >>> from sympy.combinatorics import GrayCode
        >>> a = GrayCode(3)
        >>> for i in a.generate_gray():
        ...     if i == '010':
        ...         a.skip()
        ...     print(i)
        ...
        000
        001
        011
        010
        111
        101
        100

        See Also
        ========

        generate_gray
        """
        self._skip = True

    @property
    def rank(self):
        """
        Ranks the Gray code.

        A ranking algorithm determines the position (or rank)
        of a combinatorial object among all the objects w.r.t.
        a given order. For example, the 4 bit binary reflected
        Gray code (BRGC) '0101' has a rank of 6 as it appears in
        the 6th position in the canonical ordering of the family
        of 4 bit Gray codes.

        Examples
        ========

        >>> from sympy.combinatorics import GrayCode
        >>> a = GrayCode(3)
        >>> list(a.generate_gray())
        ['000', '001', '011', '010', '110', '111', '101', '100']
        >>> GrayCode(3, start='100').rank
        7
        >>> GrayCode(3, rank=7).current
        '100'

        See Also
        ========

        unrank

        References
        ==========

        .. [1] https://web.archive.org/web/20200224064753/http://statweb.stanford.edu/~susan/courses/s208/node12.html

        """
        if self._rank is None:
            self._rank = int(gray_to_bin(self.current), 2)
        return self._rank

    @property
    def current(self):
        """
        Returns the currently referenced Gray code as a bit string.

        Examples
        ========

        >>> from sympy.combinatorics import GrayCode
        >>> GrayCode(3, start='100').current
        '100'
        """
        rv = self._current or '0'
        if not isinstance(rv, str):
            rv = bin(rv)[2:]
        return rv.rjust(self.n, '0')

    @classmethod
    def unrank(self, n, rank):
        """
        Unranks an n-bit sized Gray code of rank k. This method exists
        so that a derivative GrayCode class can define its own code of
        a given rank.

        The string here is generated in reverse order to allow for tail-call
        optimization.

        Examples
        ========

        >>> from sympy.combinatorics import GrayCode
        >>> GrayCode(5, rank=3).current
        '00010'
        >>> GrayCode.unrank(5, 3)
        '00010'

        See Also
        ========

        rank
        """
        def _unrank(k, n):
            if n == 1:
                return str(k % 2)
            m = 2**(n - 1)
            if k < m:
                return '0' + _unrank(k, n - 1)
            return '1' + _unrank(m - (k % m) - 1, n - 1)
        return _unrank(rank, n)


def random_bitstring(n):
    """
    Generates a random bitlist of length n.

    Examples
    ========

    >>> from sympy.combinatorics.graycode import random_bitstring
    >>> random_bitstring(3) # doctest: +SKIP
    100
    """
    return ''.join([random.choice('01') for i in range(n)])


def gray_to_bin(bin_list):
    """
    Convert from Gray coding to binary coding.

    We assume big endian encoding.

    Examples
    ========

    >>> from sympy.combinatorics.graycode import gray_to_bin
    >>> gray_to_bin('100')
    '111'

    See Also
    ========

    bin_to_gray
    """
    b = [bin_list[0]]
    for i in range(1, len(bin_list)):
        b += str(int(b[i - 1] != bin_list[i]))
    return ''.join(b)


def bin_to_gray(bin_list):
    """
    Convert from binary coding to gray coding.

    We assume big endian encoding.

    Examples
    ========

    >>> from sympy.combinatorics.graycode import bin_to_gray
    >>> bin_to_gray('111')
    '100'

    See Also
    ========

    gray_to_bin
    """
    b = [bin_list[0]]
    for i in range(1, len(bin_list)):
        b += str(int(bin_list[i]) ^ int(bin_list[i - 1]))
    return ''.join(b)


def get_subset_from_bitstring(super_set, bitstring):
    """
    Gets the subset defined by the bitstring.

    Examples
    ========

    >>> from sympy.combinatorics.graycode import get_subset_from_bitstring
    >>> get_subset_from_bitstring(['a', 'b', 'c', 'd'], '0011')
    ['c', 'd']
    >>> get_subset_from_bitstring(['c', 'a', 'c', 'c'], '1100')
    ['c', 'a']

    See Also
    ========

    graycode_subsets
    """
    if len(super_set) != len(bitstring):
        raise ValueError("The sizes of the lists are not equal")
    return [super_set[i] for i, j in enumerate(bitstring)
            if bitstring[i] == '1']


def graycode_subsets(gray_code_set):
    """
    Generates the subsets as enumerated by a Gray code.

    Examples
    ========

    >>> from sympy.combinatorics.graycode import graycode_subsets
    >>> list(graycode_subsets(['a', 'b', 'c']))
    [[], ['c'], ['b', 'c'], ['b'], ['a', 'b'], ['a', 'b', 'c'], \
    ['a', 'c'], ['a']]
    >>> list(graycode_subsets(['a', 'b', 'c', 'c']))
    [[], ['c'], ['c', 'c'], ['c'], ['b', 'c'], ['b', 'c', 'c'], \
    ['b', 'c'], ['b'], ['a', 'b'], ['a', 'b', 'c'], ['a', 'b', 'c', 'c'], \
    ['a', 'b', 'c'], ['a', 'c'], ['a', 'c', 'c'], ['a', 'c'], ['a']]

    See Also
    ========

    get_subset_from_bitstring
    """
    for bitstring in list(GrayCode(len(gray_code_set)).generate_gray()):
        yield get_subset_from_bitstring(gray_code_set, bitstring)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\werkzeug\formparser.py ===
from __future__ import annotations

import typing as t
from io import BytesIO
from urllib.parse import parse_qsl

from ._internal import _plain_int
from .datastructures import FileStorage
from .datastructures import Headers
from .datastructures import MultiDict
from .exceptions import RequestEntityTooLarge
from .http import parse_options_header
from .sansio.multipart import Data
from .sansio.multipart import Epilogue
from .sansio.multipart import Field
from .sansio.multipart import File
from .sansio.multipart import MultipartDecoder
from .sansio.multipart import NeedData
from .wsgi import get_content_length
from .wsgi import get_input_stream

# there are some platforms where SpooledTemporaryFile is not available.
# In that case we need to provide a fallback.
try:
    from tempfile import SpooledTemporaryFile
except ImportError:
    from tempfile import TemporaryFile

    SpooledTemporaryFile = None  # type: ignore

if t.TYPE_CHECKING:
    import typing as te

    from _typeshed.wsgi import WSGIEnvironment

    t_parse_result = tuple[
        t.IO[bytes], MultiDict[str, str], MultiDict[str, FileStorage]
    ]

    class TStreamFactory(te.Protocol):
        def __call__(
            self,
            total_content_length: int | None,
            content_type: str | None,
            filename: str | None,
            content_length: int | None = None,
        ) -> t.IO[bytes]: ...


F = t.TypeVar("F", bound=t.Callable[..., t.Any])


def default_stream_factory(
    total_content_length: int | None,
    content_type: str | None,
    filename: str | None,
    content_length: int | None = None,
) -> t.IO[bytes]:
    max_size = 1024 * 500

    if SpooledTemporaryFile is not None:
        return t.cast(t.IO[bytes], SpooledTemporaryFile(max_size=max_size, mode="rb+"))
    elif total_content_length is None or total_content_length > max_size:
        return t.cast(t.IO[bytes], TemporaryFile("rb+"))

    return BytesIO()


def parse_form_data(
    environ: WSGIEnvironment,
    stream_factory: TStreamFactory | None = None,
    max_form_memory_size: int | None = None,
    max_content_length: int | None = None,
    cls: type[MultiDict[str, t.Any]] | None = None,
    silent: bool = True,
    *,
    max_form_parts: int | None = None,
) -> t_parse_result:
    """Parse the form data in the environ and return it as tuple in the form
    ``(stream, form, files)``.  You should only call this method if the
    transport method is `POST`, `PUT`, or `PATCH`.

    If the mimetype of the data transmitted is `multipart/form-data` the
    files multidict will be filled with `FileStorage` objects.  If the
    mimetype is unknown the input stream is wrapped and returned as first
    argument, else the stream is empty.

    This is a shortcut for the common usage of :class:`FormDataParser`.

    :param environ: the WSGI environment to be used for parsing.
    :param stream_factory: An optional callable that returns a new read and
                           writeable file descriptor.  This callable works
                           the same as :meth:`Response._get_file_stream`.
    :param max_form_memory_size: the maximum number of bytes to be accepted for
                           in-memory stored form data.  If the data
                           exceeds the value specified an
                           :exc:`~exceptions.RequestEntityTooLarge`
                           exception is raised.
    :param max_content_length: If this is provided and the transmitted data
                               is longer than this value an
                               :exc:`~exceptions.RequestEntityTooLarge`
                               exception is raised.
    :param cls: an optional dict class to use.  If this is not specified
                       or `None` the default :class:`MultiDict` is used.
    :param silent: If set to False parsing errors will not be caught.
    :param max_form_parts: The maximum number of multipart parts to be parsed. If this
        is exceeded, a :exc:`~exceptions.RequestEntityTooLarge` exception is raised.
    :return: A tuple in the form ``(stream, form, files)``.

    .. versionchanged:: 3.0
        The ``charset`` and ``errors`` parameters were removed.

    .. versionchanged:: 2.3
        Added the ``max_form_parts`` parameter.

    .. versionadded:: 0.5.1
       Added the ``silent`` parameter.

    .. versionadded:: 0.5
       Added the ``max_form_memory_size``, ``max_content_length``, and ``cls``
       parameters.
    """
    return FormDataParser(
        stream_factory=stream_factory,
        max_form_memory_size=max_form_memory_size,
        max_content_length=max_content_length,
        max_form_parts=max_form_parts,
        silent=silent,
        cls=cls,
    ).parse_from_environ(environ)


class FormDataParser:
    """This class implements parsing of form data for Werkzeug.  By itself
    it can parse multipart and url encoded form data.  It can be subclassed
    and extended but for most mimetypes it is a better idea to use the
    untouched stream and expose it as separate attributes on a request
    object.

    :param stream_factory: An optional callable that returns a new read and
                           writeable file descriptor.  This callable works
                           the same as :meth:`Response._get_file_stream`.
    :param max_form_memory_size: the maximum number of bytes to be accepted for
                           in-memory stored form data.  If the data
                           exceeds the value specified an
                           :exc:`~exceptions.RequestEntityTooLarge`
                           exception is raised.
    :param max_content_length: If this is provided and the transmitted data
                               is longer than this value an
                               :exc:`~exceptions.RequestEntityTooLarge`
                               exception is raised.
    :param cls: an optional dict class to use.  If this is not specified
                       or `None` the default :class:`MultiDict` is used.
    :param silent: If set to False parsing errors will not be caught.
    :param max_form_parts: The maximum number of multipart parts to be parsed. If this
        is exceeded, a :exc:`~exceptions.RequestEntityTooLarge` exception is raised.

    .. versionchanged:: 3.0
        The ``charset`` and ``errors`` parameters were removed.

    .. versionchanged:: 3.0
        The ``parse_functions`` attribute and ``get_parse_func`` methods were removed.

    .. versionchanged:: 2.2.3
        Added the ``max_form_parts`` parameter.

    .. versionadded:: 0.8
    """

    def __init__(
        self,
        stream_factory: TStreamFactory | None = None,
        max_form_memory_size: int | None = None,
        max_content_length: int | None = None,
        cls: type[MultiDict[str, t.Any]] | None = None,
        silent: bool = True,
        *,
        max_form_parts: int | None = None,
    ) -> None:
        if stream_factory is None:
            stream_factory = default_stream_factory

        self.stream_factory = stream_factory
        self.max_form_memory_size = max_form_memory_size
        self.max_content_length = max_content_length
        self.max_form_parts = max_form_parts

        if cls is None:
            cls = t.cast("type[MultiDict[str, t.Any]]", MultiDict)

        self.cls = cls
        self.silent = silent

    def parse_from_environ(self, environ: WSGIEnvironment) -> t_parse_result:
        """Parses the information from the environment as form data.

        :param environ: the WSGI environment to be used for parsing.
        :return: A tuple in the form ``(stream, form, files)``.
        """
        stream = get_input_stream(environ, max_content_length=self.max_content_length)
        content_length = get_content_length(environ)
        mimetype, options = parse_options_header(environ.get("CONTENT_TYPE"))
        return self.parse(
            stream,
            content_length=content_length,
            mimetype=mimetype,
            options=options,
        )

    def parse(
        self,
        stream: t.IO[bytes],
        mimetype: str,
        content_length: int | None,
        options: dict[str, str] | None = None,
    ) -> t_parse_result:
        """Parses the information from the given stream, mimetype,
        content length and mimetype parameters.

        :param stream: an input stream
        :param mimetype: the mimetype of the data
        :param content_length: the content length of the incoming data
        :param options: optional mimetype parameters (used for
                        the multipart boundary for instance)
        :return: A tuple in the form ``(stream, form, files)``.

        .. versionchanged:: 3.0
            The invalid ``application/x-url-encoded`` content type is not
            treated as ``application/x-www-form-urlencoded``.
        """
        if mimetype == "multipart/form-data":
            parse_func = self._parse_multipart
        elif mimetype == "application/x-www-form-urlencoded":
            parse_func = self._parse_urlencoded
        else:
            return stream, self.cls(), self.cls()

        if options is None:
            options = {}

        try:
            return parse_func(stream, mimetype, content_length, options)
        except ValueError:
            if not self.silent:
                raise

        return stream, self.cls(), self.cls()

    def _parse_multipart(
        self,
        stream: t.IO[bytes],
        mimetype: str,
        content_length: int | None,
        options: dict[str, str],
    ) -> t_parse_result:
        parser = MultiPartParser(
            stream_factory=self.stream_factory,
            max_form_memory_size=self.max_form_memory_size,
            max_form_parts=self.max_form_parts,
            cls=self.cls,
        )
        boundary = options.get("boundary", "").encode("ascii")

        if not boundary:
            raise ValueError("Missing boundary")

        form, files = parser.parse(stream, boundary, content_length)
        return stream, form, files

    def _parse_urlencoded(
        self,
        stream: t.IO[bytes],
        mimetype: str,
        content_length: int | None,
        options: dict[str, str],
    ) -> t_parse_result:
        if (
            self.max_form_memory_size is not None
            and content_length is not None
            and content_length > self.max_form_memory_size
        ):
            raise RequestEntityTooLarge()

        items = parse_qsl(
            stream.read().decode(),
            keep_blank_values=True,
            errors="werkzeug.url_quote",
        )
        return stream, self.cls(items), self.cls()


class MultiPartParser:
    def __init__(
        self,
        stream_factory: TStreamFactory | None = None,
        max_form_memory_size: int | None = None,
        cls: type[MultiDict[str, t.Any]] | None = None,
        buffer_size: int = 64 * 1024,
        max_form_parts: int | None = None,
    ) -> None:
        self.max_form_memory_size = max_form_memory_size
        self.max_form_parts = max_form_parts

        if stream_factory is None:
            stream_factory = default_stream_factory

        self.stream_factory = stream_factory

        if cls is None:
            cls = t.cast("type[MultiDict[str, t.Any]]", MultiDict)

        self.cls = cls
        self.buffer_size = buffer_size

    def fail(self, message: str) -> te.NoReturn:
        raise ValueError(message)

    def get_part_charset(self, headers: Headers) -> str:
        # Figure out input charset for current part
        content_type = headers.get("content-type")

        if content_type:
            parameters = parse_options_header(content_type)[1]
            ct_charset = parameters.get("charset", "").lower()

            # A safe list of encodings. Modern clients should only send ASCII or UTF-8.
            # This list will not be extended further.
            if ct_charset in {"ascii", "us-ascii", "utf-8", "iso-8859-1"}:
                return ct_charset

        return "utf-8"

    def start_file_streaming(
        self, event: File, total_content_length: int | None
    ) -> t.IO[bytes]:
        content_type = event.headers.get("content-type")

        try:
            content_length = _plain_int(event.headers["content-length"])
        except (KeyError, ValueError):
            content_length = 0

        container = self.stream_factory(
            total_content_length=total_content_length,
            filename=event.filename,
            content_type=content_type,
            content_length=content_length,
        )
        return container

    def parse(
        self, stream: t.IO[bytes], boundary: bytes, content_length: int | None
    ) -> tuple[MultiDict[str, str], MultiDict[str, FileStorage]]:
        current_part: Field | File
        field_size: int | None = None
        container: t.IO[bytes] | list[bytes]
        _write: t.Callable[[bytes], t.Any]

        parser = MultipartDecoder(
            boundary,
            max_form_memory_size=self.max_form_memory_size,
            max_parts=self.max_form_parts,
        )

        fields = []
        files = []

        for data in _chunk_iter(stream.read, self.buffer_size):
            parser.receive_data(data)
            event = parser.next_event()
            while not isinstance(event, (Epilogue, NeedData)):
                if isinstance(event, Field):
                    current_part = event
                    field_size = 0
                    container = []
                    _write = container.append
                elif isinstance(event, File):
                    current_part = event
                    field_size = None
                    container = self.start_file_streaming(event, content_length)
                    _write = container.write
                elif isinstance(event, Data):
                    if self.max_form_memory_size is not None and field_size is not None:
                        # Ensure that accumulated data events do not exceed limit.
                        # Also checked within single event in MultipartDecoder.
                        field_size += len(event.data)

                        if field_size > self.max_form_memory_size:
                            raise RequestEntityTooLarge()

                    _write(event.data)
                    if not event.more_data:
                        if isinstance(current_part, Field):
                            value = b"".join(container).decode(
                                self.get_part_charset(current_part.headers), "replace"
                            )
                            fields.append((current_part.name, value))
                        else:
                            container = t.cast(t.IO[bytes], container)
                            container.seek(0)
                            files.append(
                                (
                                    current_part.name,
                                    FileStorage(
                                        container,
                                        current_part.filename,
                                        current_part.name,
                                        headers=current_part.headers,
                                    ),
                                )
                            )

                event = parser.next_event()

        return self.cls(fields), self.cls(files)


def _chunk_iter(read: t.Callable[[int], bytes], size: int) -> t.Iterator[bytes | None]:
    """Read data in chunks for multipart/form-data parsing. Stop if no data is read.
    Yield ``None`` at the end to signal end of parsing.
    """
    while True:
        data = read(size)

        if not data:
            break

        yield data

    yield None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\ext\orderinglist.py ===
# ext/orderinglist.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors

"""A custom list that manages index/position information for contained
elements.

:author: Jason Kirtland

``orderinglist`` is a helper for mutable ordered relationships.  It will
intercept list operations performed on a :func:`_orm.relationship`-managed
collection and
automatically synchronize changes in list position onto a target scalar
attribute.

Example: A ``slide`` table, where each row refers to zero or more entries
in a related ``bullet`` table.   The bullets within a slide are
displayed in order based on the value of the ``position`` column in the
``bullet`` table.   As entries are reordered in memory, the value of the
``position`` attribute should be updated to reflect the new sort order::


    Base = declarative_base()


    class Slide(Base):
        __tablename__ = "slide"

        id = Column(Integer, primary_key=True)
        name = Column(String)

        bullets = relationship("Bullet", order_by="Bullet.position")


    class Bullet(Base):
        __tablename__ = "bullet"
        id = Column(Integer, primary_key=True)
        slide_id = Column(Integer, ForeignKey("slide.id"))
        position = Column(Integer)
        text = Column(String)

The standard relationship mapping will produce a list-like attribute on each
``Slide`` containing all related ``Bullet`` objects,
but coping with changes in ordering is not handled automatically.
When appending a ``Bullet`` into ``Slide.bullets``, the ``Bullet.position``
attribute will remain unset until manually assigned.   When the ``Bullet``
is inserted into the middle of the list, the following ``Bullet`` objects
will also need to be renumbered.

The :class:`.OrderingList` object automates this task, managing the
``position`` attribute on all ``Bullet`` objects in the collection.  It is
constructed using the :func:`.ordering_list` factory::

    from sqlalchemy.ext.orderinglist import ordering_list

    Base = declarative_base()


    class Slide(Base):
        __tablename__ = "slide"

        id = Column(Integer, primary_key=True)
        name = Column(String)

        bullets = relationship(
            "Bullet",
            order_by="Bullet.position",
            collection_class=ordering_list("position"),
        )


    class Bullet(Base):
        __tablename__ = "bullet"
        id = Column(Integer, primary_key=True)
        slide_id = Column(Integer, ForeignKey("slide.id"))
        position = Column(Integer)
        text = Column(String)

With the above mapping the ``Bullet.position`` attribute is managed::

    s = Slide()
    s.bullets.append(Bullet())
    s.bullets.append(Bullet())
    s.bullets[1].position
    >>> 1
    s.bullets.insert(1, Bullet())
    s.bullets[2].position
    >>> 2

The :class:`.OrderingList` construct only works with **changes** to a
collection, and not the initial load from the database, and requires that the
list be sorted when loaded.  Therefore, be sure to specify ``order_by`` on the
:func:`_orm.relationship` against the target ordering attribute, so that the
ordering is correct when first loaded.

.. warning::

  :class:`.OrderingList` only provides limited functionality when a primary
  key column or unique column is the target of the sort.  Operations
  that are unsupported or are problematic include:

    * two entries must trade values.  This is not supported directly in the
      case of a primary key or unique constraint because it means at least
      one row would need to be temporarily removed first, or changed to
      a third, neutral value while the switch occurs.

    * an entry must be deleted in order to make room for a new entry.
      SQLAlchemy's unit of work performs all INSERTs before DELETEs within a
      single flush.  In the case of a primary key, it will trade
      an INSERT/DELETE of the same primary key for an UPDATE statement in order
      to lessen the impact of this limitation, however this does not take place
      for a UNIQUE column.
      A future feature will allow the "DELETE before INSERT" behavior to be
      possible, alleviating this limitation, though this feature will require
      explicit configuration at the mapper level for sets of columns that
      are to be handled in this way.

:func:`.ordering_list` takes the name of the related object's ordering
attribute as an argument.  By default, the zero-based integer index of the
object's position in the :func:`.ordering_list` is synchronized with the
ordering attribute: index 0 will get position 0, index 1 position 1, etc.  To
start numbering at 1 or some other integer, provide ``count_from=1``.


"""
from __future__ import annotations

from typing import Callable
from typing import List
from typing import Optional
from typing import Sequence
from typing import TypeVar

from ..orm.collections import collection
from ..orm.collections import collection_adapter

_T = TypeVar("_T")
OrderingFunc = Callable[[int, Sequence[_T]], int]


__all__ = ["ordering_list"]


def ordering_list(
    attr: str,
    count_from: Optional[int] = None,
    ordering_func: Optional[OrderingFunc] = None,
    reorder_on_append: bool = False,
) -> Callable[[], OrderingList]:
    """Prepares an :class:`OrderingList` factory for use in mapper definitions.

    Returns an object suitable for use as an argument to a Mapper
    relationship's ``collection_class`` option.  e.g.::

        from sqlalchemy.ext.orderinglist import ordering_list


        class Slide(Base):
            __tablename__ = "slide"

            id = Column(Integer, primary_key=True)
            name = Column(String)

            bullets = relationship(
                "Bullet",
                order_by="Bullet.position",
                collection_class=ordering_list("position"),
            )

    :param attr:
      Name of the mapped attribute to use for storage and retrieval of
      ordering information

    :param count_from:
      Set up an integer-based ordering, starting at ``count_from``.  For
      example, ``ordering_list('pos', count_from=1)`` would create a 1-based
      list in SQL, storing the value in the 'pos' column.  Ignored if
      ``ordering_func`` is supplied.

    Additional arguments are passed to the :class:`.OrderingList` constructor.

    """

    kw = _unsugar_count_from(
        count_from=count_from,
        ordering_func=ordering_func,
        reorder_on_append=reorder_on_append,
    )
    return lambda: OrderingList(attr, **kw)


# Ordering utility functions


def count_from_0(index, collection):
    """Numbering function: consecutive integers starting at 0."""

    return index


def count_from_1(index, collection):
    """Numbering function: consecutive integers starting at 1."""

    return index + 1


def count_from_n_factory(start):
    """Numbering function: consecutive integers starting at arbitrary start."""

    def f(index, collection):
        return index + start

    try:
        f.__name__ = "count_from_%i" % start
    except TypeError:
        pass
    return f


def _unsugar_count_from(**kw):
    """Builds counting functions from keyword arguments.

    Keyword argument filter, prepares a simple ``ordering_func`` from a
    ``count_from`` argument, otherwise passes ``ordering_func`` on unchanged.
    """

    count_from = kw.pop("count_from", None)
    if kw.get("ordering_func", None) is None and count_from is not None:
        if count_from == 0:
            kw["ordering_func"] = count_from_0
        elif count_from == 1:
            kw["ordering_func"] = count_from_1
        else:
            kw["ordering_func"] = count_from_n_factory(count_from)
    return kw


class OrderingList(List[_T]):
    """A custom list that manages position information for its children.

    The :class:`.OrderingList` object is normally set up using the
    :func:`.ordering_list` factory function, used in conjunction with
    the :func:`_orm.relationship` function.

    """

    ordering_attr: str
    ordering_func: OrderingFunc
    reorder_on_append: bool

    def __init__(
        self,
        ordering_attr: Optional[str] = None,
        ordering_func: Optional[OrderingFunc] = None,
        reorder_on_append: bool = False,
    ):
        """A custom list that manages position information for its children.

        ``OrderingList`` is a ``collection_class`` list implementation that
        syncs position in a Python list with a position attribute on the
        mapped objects.

        This implementation relies on the list starting in the proper order,
        so be **sure** to put an ``order_by`` on your relationship.

        :param ordering_attr:
          Name of the attribute that stores the object's order in the
          relationship.

        :param ordering_func: Optional.  A function that maps the position in
          the Python list to a value to store in the
          ``ordering_attr``.  Values returned are usually (but need not be!)
          integers.

          An ``ordering_func`` is called with two positional parameters: the
          index of the element in the list, and the list itself.

          If omitted, Python list indexes are used for the attribute values.
          Two basic pre-built numbering functions are provided in this module:
          ``count_from_0`` and ``count_from_1``.  For more exotic examples
          like stepped numbering, alphabetical and Fibonacci numbering, see
          the unit tests.

        :param reorder_on_append:
          Default False.  When appending an object with an existing (non-None)
          ordering value, that value will be left untouched unless
          ``reorder_on_append`` is true.  This is an optimization to avoid a
          variety of dangerous unexpected database writes.

          SQLAlchemy will add instances to the list via append() when your
          object loads.  If for some reason the result set from the database
          skips a step in the ordering (say, row '1' is missing but you get
          '2', '3', and '4'), reorder_on_append=True would immediately
          renumber the items to '1', '2', '3'.  If you have multiple sessions
          making changes, any of whom happen to load this collection even in
          passing, all of the sessions would try to "clean up" the numbering
          in their commits, possibly causing all but one to fail with a
          concurrent modification error.

          Recommend leaving this with the default of False, and just call
          ``reorder()`` if you're doing ``append()`` operations with
          previously ordered instances or when doing some housekeeping after
          manual sql operations.

        """
        self.ordering_attr = ordering_attr
        if ordering_func is None:
            ordering_func = count_from_0
        self.ordering_func = ordering_func
        self.reorder_on_append = reorder_on_append

    # More complex serialization schemes (multi column, e.g.) are possible by
    # subclassing and reimplementing these two methods.
    def _get_order_value(self, entity):
        return getattr(entity, self.ordering_attr)

    def _set_order_value(self, entity, value):
        setattr(entity, self.ordering_attr, value)

    def reorder(self) -> None:
        """Synchronize ordering for the entire collection.

        Sweeps through the list and ensures that each object has accurate
        ordering information set.

        """
        for index, entity in enumerate(self):
            self._order_entity(index, entity, True)

    # As of 0.5, _reorder is no longer semi-private
    _reorder = reorder

    def _order_entity(self, index, entity, reorder=True):
        have = self._get_order_value(entity)

        # Don't disturb existing ordering if reorder is False
        if have is not None and not reorder:
            return

        should_be = self.ordering_func(index, self)
        if have != should_be:
            self._set_order_value(entity, should_be)

    def append(self, entity):
        super().append(entity)
        self._order_entity(len(self) - 1, entity, self.reorder_on_append)

    def _raw_append(self, entity):
        """Append without any ordering behavior."""

        super().append(entity)

    _raw_append = collection.adds(1)(_raw_append)

    def insert(self, index, entity):
        super().insert(index, entity)
        self._reorder()

    def remove(self, entity):
        super().remove(entity)

        adapter = collection_adapter(self)
        if adapter and adapter._referenced_by_owner:
            self._reorder()

    def pop(self, index=-1):
        entity = super().pop(index)
        self._reorder()
        return entity

    def __setitem__(self, index, entity):
        if isinstance(index, slice):
            step = index.step or 1
            start = index.start or 0
            if start < 0:
                start += len(self)
            stop = index.stop or len(self)
            if stop < 0:
                stop += len(self)

            for i in range(start, stop, step):
                self.__setitem__(i, entity[i])
        else:
            self._order_entity(index, entity, True)
            super().__setitem__(index, entity)

    def __delitem__(self, index):
        super().__delitem__(index)
        self._reorder()

    def __setslice__(self, start, end, values):
        super().__setslice__(start, end, values)
        self._reorder()

    def __delslice__(self, start, end):
        super().__delslice__(start, end)
        self._reorder()

    def __reduce__(self):
        return _reconstitute, (self.__class__, self.__dict__, list(self))

    for func_name, func in list(locals().items()):
        if (
            callable(func)
            and func.__name__ == func_name
            and not func.__doc__
            and hasattr(list, func_name)
        ):
            func.__doc__ = getattr(list, func_name).__doc__
    del func_name, func


def _reconstitute(cls, dict_, items):
    """Reconstitute an :class:`.OrderingList`.

    This is the adjoint to :meth:`.OrderingList.__reduce__`.  It is used for
    unpickling :class:`.OrderingList` objects.

    """
    obj = cls.__new__(cls)
    obj.__dict__.update(dict_)
    list.extend(obj, items)
    return obj