
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\arrays\sparse\test_reductions.py ===
import numpy as np
import pytest

from pandas import (
    NaT,
    SparseDtype,
    Timestamp,
    isna,
)
from pandas.core.arrays.sparse import SparseArray


class TestReductions:
    @pytest.mark.parametrize(
        "data,pos,neg",
        [
            ([True, True, True], True, False),
            ([1, 2, 1], 1, 0),
            ([1.0, 2.0, 1.0], 1.0, 0.0),
        ],
    )
    def test_all(self, data, pos, neg):
        # GH#17570
        out = SparseArray(data).all()
        assert out

        out = SparseArray(data, fill_value=pos).all()
        assert out

        data[1] = neg
        out = SparseArray(data).all()
        assert not out

        out = SparseArray(data, fill_value=pos).all()
        assert not out

    @pytest.mark.parametrize(
        "data,pos,neg",
        [
            ([True, True, True], True, False),
            ([1, 2, 1], 1, 0),
            ([1.0, 2.0, 1.0], 1.0, 0.0),
        ],
    )
    def test_numpy_all(self, data, pos, neg):
        # GH#17570
        out = np.all(SparseArray(data))
        assert out

        out = np.all(SparseArray(data, fill_value=pos))
        assert out

        data[1] = neg
        out = np.all(SparseArray(data))
        assert not out

        out = np.all(SparseArray(data, fill_value=pos))
        assert not out

        # raises with a different message on py2.
        msg = "the 'out' parameter is not supported"
        with pytest.raises(ValueError, match=msg):
            np.all(SparseArray(data), out=np.array([]))

    @pytest.mark.parametrize(
        "data,pos,neg",
        [
            ([False, True, False], True, False),
            ([0, 2, 0], 2, 0),
            ([0.0, 2.0, 0.0], 2.0, 0.0),
        ],
    )
    def test_any(self, data, pos, neg):
        # GH#17570
        out = SparseArray(data).any()
        assert out

        out = SparseArray(data, fill_value=pos).any()
        assert out

        data[1] = neg
        out = SparseArray(data).any()
        assert not out

        out = SparseArray(data, fill_value=pos).any()
        assert not out

    @pytest.mark.parametrize(
        "data,pos,neg",
        [
            ([False, True, False], True, False),
            ([0, 2, 0], 2, 0),
            ([0.0, 2.0, 0.0], 2.0, 0.0),
        ],
    )
    def test_numpy_any(self, data, pos, neg):
        # GH#17570
        out = np.any(SparseArray(data))
        assert out

        out = np.any(SparseArray(data, fill_value=pos))
        assert out

        data[1] = neg
        out = np.any(SparseArray(data))
        assert not out

        out = np.any(SparseArray(data, fill_value=pos))
        assert not out

        msg = "the 'out' parameter is not supported"
        with pytest.raises(ValueError, match=msg):
            np.any(SparseArray(data), out=out)

    def test_sum(self):
        data = np.arange(10).astype(float)
        out = SparseArray(data).sum()
        assert out == 45.0

        data[5] = np.nan
        out = SparseArray(data, fill_value=2).sum()
        assert out == 40.0

        out = SparseArray(data, fill_value=np.nan).sum()
        assert out == 40.0

    @pytest.mark.parametrize(
        "arr",
        [np.array([0, 1, np.nan, 1]), np.array([0, 1, 1])],
    )
    @pytest.mark.parametrize("fill_value", [0, 1, np.nan])
    @pytest.mark.parametrize("min_count, expected", [(3, 2), (4, np.nan)])
    def test_sum_min_count(self, arr, fill_value, min_count, expected):
        # GH#25777
        sparray = SparseArray(arr, fill_value=fill_value)
        result = sparray.sum(min_count=min_count)
        if np.isnan(expected):
            assert np.isnan(result)
        else:
            assert result == expected

    def test_bool_sum_min_count(self):
        spar_bool = SparseArray([False, True] * 5, dtype=np.bool_, fill_value=True)
        res = spar_bool.sum(min_count=1)
        assert res == 5
        res = spar_bool.sum(min_count=11)
        assert isna(res)

    def test_numpy_sum(self):
        data = np.arange(10).astype(float)
        out = np.sum(SparseArray(data))
        assert out == 45.0

        data[5] = np.nan
        out = np.sum(SparseArray(data, fill_value=2))
        assert out == 40.0

        out = np.sum(SparseArray(data, fill_value=np.nan))
        assert out == 40.0

        msg = "the 'dtype' parameter is not supported"
        with pytest.raises(ValueError, match=msg):
            np.sum(SparseArray(data), dtype=np.int64)

        msg = "the 'out' parameter is not supported"
        with pytest.raises(ValueError, match=msg):
            np.sum(SparseArray(data), out=out)

    def test_mean(self):
        data = np.arange(10).astype(float)
        out = SparseArray(data).mean()
        assert out == 4.5

        data[5] = np.nan
        out = SparseArray(data).mean()
        assert out == 40.0 / 9

    def test_numpy_mean(self):
        data = np.arange(10).astype(float)
        out = np.mean(SparseArray(data))
        assert out == 4.5

        data[5] = np.nan
        out = np.mean(SparseArray(data))
        assert out == 40.0 / 9

        msg = "the 'dtype' parameter is not supported"
        with pytest.raises(ValueError, match=msg):
            np.mean(SparseArray(data), dtype=np.int64)

        msg = "the 'out' parameter is not supported"
        with pytest.raises(ValueError, match=msg):
            np.mean(SparseArray(data), out=out)


class TestMinMax:
    @pytest.mark.parametrize(
        "raw_data,max_expected,min_expected",
        [
            (np.arange(5.0), [4], [0]),
            (-np.arange(5.0), [0], [-4]),
            (np.array([0, 1, 2, np.nan, 4]), [4], [0]),
            (np.array([np.nan] * 5), [np.nan], [np.nan]),
            (np.array([]), [np.nan], [np.nan]),
        ],
    )
    def test_nan_fill_value(self, raw_data, max_expected, min_expected):
        arr = SparseArray(raw_data)
        max_result = arr.max()
        min_result = arr.min()
        assert max_result in max_expected
        assert min_result in min_expected

        max_result = arr.max(skipna=False)
        min_result = arr.min(skipna=False)
        if np.isnan(raw_data).any():
            assert np.isnan(max_result)
            assert np.isnan(min_result)
        else:
            assert max_result in max_expected
            assert min_result in min_expected

    @pytest.mark.parametrize(
        "fill_value,max_expected,min_expected",
        [
            (100, 100, 0),
            (-100, 1, -100),
        ],
    )
    def test_fill_value(self, fill_value, max_expected, min_expected):
        arr = SparseArray(
            np.array([fill_value, 0, 1]), dtype=SparseDtype("int", fill_value)
        )
        max_result = arr.max()
        assert max_result == max_expected

        min_result = arr.min()
        assert min_result == min_expected

    def test_only_fill_value(self):
        fv = 100
        arr = SparseArray(np.array([fv, fv, fv]), dtype=SparseDtype("int", fv))
        assert len(arr._valid_sp_values) == 0

        assert arr.max() == fv
        assert arr.min() == fv
        assert arr.max(skipna=False) == fv
        assert arr.min(skipna=False) == fv

    @pytest.mark.parametrize("func", ["min", "max"])
    @pytest.mark.parametrize("data", [np.array([]), np.array([np.nan, np.nan])])
    @pytest.mark.parametrize(
        "dtype,expected",
        [
            (SparseDtype(np.float64, np.nan), np.nan),
            (SparseDtype(np.float64, 5.0), np.nan),
            (SparseDtype("datetime64[ns]", NaT), NaT),
            (SparseDtype("datetime64[ns]", Timestamp("2018-05-05")), NaT),
        ],
    )
    def test_na_value_if_no_valid_values(self, func, data, dtype, expected):
        arr = SparseArray(data, dtype=dtype)
        result = getattr(arr, func)()
        if expected is NaT:
            # TODO: pin down whether we wrap datetime64("NaT")
            assert result is NaT or np.isnat(result)
        else:
            assert np.isnan(result)


class TestArgmaxArgmin:
    @pytest.mark.parametrize(
        "arr,argmax_expected,argmin_expected",
        [
            (SparseArray([1, 2, 0, 1, 2]), 1, 2),
            (SparseArray([-1, -2, 0, -1, -2]), 2, 1),
            (SparseArray([np.nan, 1, 0, 0, np.nan, -1]), 1, 5),
            (SparseArray([np.nan, 1, 0, 0, np.nan, 2]), 5, 2),
            (SparseArray([np.nan, 1, 0, 0, np.nan, 2], fill_value=-1), 5, 2),
            (SparseArray([np.nan, 1, 0, 0, np.nan, 2], fill_value=0), 5, 2),
            (SparseArray([np.nan, 1, 0, 0, np.nan, 2], fill_value=1), 5, 2),
            (SparseArray([np.nan, 1, 0, 0, np.nan, 2], fill_value=2), 5, 2),
            (SparseArray([np.nan, 1, 0, 0, np.nan, 2], fill_value=3), 5, 2),
            (SparseArray([0] * 10 + [-1], fill_value=0), 0, 10),
            (SparseArray([0] * 10 + [-1], fill_value=-1), 0, 10),
            (SparseArray([0] * 10 + [-1], fill_value=1), 0, 10),
            (SparseArray([-1] + [0] * 10, fill_value=0), 1, 0),
            (SparseArray([1] + [0] * 10, fill_value=0), 0, 1),
            (SparseArray([-1] + [0] * 10, fill_value=-1), 1, 0),
            (SparseArray([1] + [0] * 10, fill_value=1), 0, 1),
        ],
    )
    def test_argmax_argmin(self, arr, argmax_expected, argmin_expected):
        argmax_result = arr.argmax()
        argmin_result = arr.argmin()
        assert argmax_result == argmax_expected
        assert argmin_result == argmin_expected

    @pytest.mark.parametrize(
        "arr,method",
        [(SparseArray([]), "argmax"), (SparseArray([]), "argmin")],
    )
    def test_empty_array(self, arr, method):
        msg = f"attempt to get {method} of an empty sequence"
        with pytest.raises(ValueError, match=msg):
            arr.argmax() if method == "argmax" else arr.argmin()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\frame\methods\test_compare.py ===
import numpy as np
import pytest

from pandas.compat.numpy import np_version_gte1p25

import pandas as pd
import pandas._testing as tm


@pytest.mark.parametrize("align_axis", [0, 1, "index", "columns"])
def test_compare_axis(align_axis):
    # GH#30429
    df = pd.DataFrame(
        {"col1": ["a", "b", "c"], "col2": [1.0, 2.0, np.nan], "col3": [1.0, 2.0, 3.0]},
        columns=["col1", "col2", "col3"],
    )
    df2 = df.copy()
    df2.loc[0, "col1"] = "c"
    df2.loc[2, "col3"] = 4.0

    result = df.compare(df2, align_axis=align_axis)

    if align_axis in (1, "columns"):
        indices = pd.Index([0, 2])
        columns = pd.MultiIndex.from_product([["col1", "col3"], ["self", "other"]])
        expected = pd.DataFrame(
            [["a", "c", np.nan, np.nan], [np.nan, np.nan, 3.0, 4.0]],
            index=indices,
            columns=columns,
        )
    else:
        indices = pd.MultiIndex.from_product([[0, 2], ["self", "other"]])
        columns = pd.Index(["col1", "col3"])
        expected = pd.DataFrame(
            [["a", np.nan], ["c", np.nan], [np.nan, 3.0], [np.nan, 4.0]],
            index=indices,
            columns=columns,
        )
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "keep_shape, keep_equal",
    [
        (True, False),
        (False, True),
        (True, True),
        # False, False case is already covered in test_compare_axis
    ],
)
def test_compare_various_formats(keep_shape, keep_equal):
    df = pd.DataFrame(
        {"col1": ["a", "b", "c"], "col2": [1.0, 2.0, np.nan], "col3": [1.0, 2.0, 3.0]},
        columns=["col1", "col2", "col3"],
    )
    df2 = df.copy()
    df2.loc[0, "col1"] = "c"
    df2.loc[2, "col3"] = 4.0

    result = df.compare(df2, keep_shape=keep_shape, keep_equal=keep_equal)

    if keep_shape:
        indices = pd.Index([0, 1, 2])
        columns = pd.MultiIndex.from_product(
            [["col1", "col2", "col3"], ["self", "other"]]
        )
        if keep_equal:
            expected = pd.DataFrame(
                [
                    ["a", "c", 1.0, 1.0, 1.0, 1.0],
                    ["b", "b", 2.0, 2.0, 2.0, 2.0],
                    ["c", "c", np.nan, np.nan, 3.0, 4.0],
                ],
                index=indices,
                columns=columns,
            )
        else:
            expected = pd.DataFrame(
                [
                    ["a", "c", np.nan, np.nan, np.nan, np.nan],
                    [np.nan, np.nan, np.nan, np.nan, np.nan, np.nan],
                    [np.nan, np.nan, np.nan, np.nan, 3.0, 4.0],
                ],
                index=indices,
                columns=columns,
            )
    else:
        indices = pd.Index([0, 2])
        columns = pd.MultiIndex.from_product([["col1", "col3"], ["self", "other"]])
        expected = pd.DataFrame(
            [["a", "c", 1.0, 1.0], ["c", "c", 3.0, 4.0]], index=indices, columns=columns
        )
    tm.assert_frame_equal(result, expected)


def test_compare_with_equal_nulls():
    # We want to make sure two NaNs are considered the same
    # and dropped where applicable
    df = pd.DataFrame(
        {"col1": ["a", "b", "c"], "col2": [1.0, 2.0, np.nan], "col3": [1.0, 2.0, 3.0]},
        columns=["col1", "col2", "col3"],
    )
    df2 = df.copy()
    df2.loc[0, "col1"] = "c"

    result = df.compare(df2)
    indices = pd.Index([0])
    columns = pd.MultiIndex.from_product([["col1"], ["self", "other"]])
    expected = pd.DataFrame([["a", "c"]], index=indices, columns=columns)
    tm.assert_frame_equal(result, expected)


def test_compare_with_non_equal_nulls():
    # We want to make sure the relevant NaNs do not get dropped
    # even if the entire row or column are NaNs
    df = pd.DataFrame(
        {"col1": ["a", "b", "c"], "col2": [1.0, 2.0, np.nan], "col3": [1.0, 2.0, 3.0]},
        columns=["col1", "col2", "col3"],
    )
    df2 = df.copy()
    df2.loc[0, "col1"] = "c"
    df2.loc[2, "col3"] = np.nan

    result = df.compare(df2)

    indices = pd.Index([0, 2])
    columns = pd.MultiIndex.from_product([["col1", "col3"], ["self", "other"]])
    expected = pd.DataFrame(
        [["a", "c", np.nan, np.nan], [np.nan, np.nan, 3.0, np.nan]],
        index=indices,
        columns=columns,
    )
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("align_axis", [0, 1])
def test_compare_multi_index(align_axis):
    df = pd.DataFrame(
        {"col1": ["a", "b", "c"], "col2": [1.0, 2.0, np.nan], "col3": [1.0, 2.0, 3.0]}
    )
    df.columns = pd.MultiIndex.from_arrays([["a", "a", "b"], ["col1", "col2", "col3"]])
    df.index = pd.MultiIndex.from_arrays([["x", "x", "y"], [0, 1, 2]])

    df2 = df.copy()
    df2.iloc[0, 0] = "c"
    df2.iloc[2, 2] = 4.0

    result = df.compare(df2, align_axis=align_axis)

    if align_axis == 0:
        indices = pd.MultiIndex.from_arrays(
            [["x", "x", "y", "y"], [0, 0, 2, 2], ["self", "other", "self", "other"]]
        )
        columns = pd.MultiIndex.from_arrays([["a", "b"], ["col1", "col3"]])
        data = [["a", np.nan], ["c", np.nan], [np.nan, 3.0], [np.nan, 4.0]]
    else:
        indices = pd.MultiIndex.from_arrays([["x", "y"], [0, 2]])
        columns = pd.MultiIndex.from_arrays(
            [
                ["a", "a", "b", "b"],
                ["col1", "col1", "col3", "col3"],
                ["self", "other", "self", "other"],
            ]
        )
        data = [["a", "c", np.nan, np.nan], [np.nan, np.nan, 3.0, 4.0]]

    expected = pd.DataFrame(data=data, index=indices, columns=columns)
    tm.assert_frame_equal(result, expected)


def test_compare_unaligned_objects():
    # test DataFrames with different indices
    msg = (
        r"Can only compare identically-labeled \(both index and columns\) DataFrame "
        "objects"
    )
    with pytest.raises(ValueError, match=msg):
        df1 = pd.DataFrame([1, 2, 3], index=["a", "b", "c"])
        df2 = pd.DataFrame([1, 2, 3], index=["a", "b", "d"])
        df1.compare(df2)

    # test DataFrames with different shapes
    msg = (
        r"Can only compare identically-labeled \(both index and columns\) DataFrame "
        "objects"
    )
    with pytest.raises(ValueError, match=msg):
        df1 = pd.DataFrame(np.ones((3, 3)))
        df2 = pd.DataFrame(np.zeros((2, 1)))
        df1.compare(df2)


def test_compare_result_names():
    # GH 44354
    df1 = pd.DataFrame(
        {"col1": ["a", "b", "c"], "col2": [1.0, 2.0, np.nan], "col3": [1.0, 2.0, 3.0]},
    )
    df2 = pd.DataFrame(
        {
            "col1": ["c", "b", "c"],
            "col2": [1.0, 2.0, np.nan],
            "col3": [1.0, 2.0, np.nan],
        },
    )
    result = df1.compare(df2, result_names=("left", "right"))
    expected = pd.DataFrame(
        {
            ("col1", "left"): {0: "a", 2: np.nan},
            ("col1", "right"): {0: "c", 2: np.nan},
            ("col3", "left"): {0: np.nan, 2: 3.0},
            ("col3", "right"): {0: np.nan, 2: np.nan},
        }
    )
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "result_names",
    [
        [1, 2],
        "HK",
        {"2": 2, "3": 3},
        3,
        3.0,
    ],
)
def test_invalid_input_result_names(result_names):
    # GH 44354
    df1 = pd.DataFrame(
        {"col1": ["a", "b", "c"], "col2": [1.0, 2.0, np.nan], "col3": [1.0, 2.0, 3.0]},
    )
    df2 = pd.DataFrame(
        {
            "col1": ["c", "b", "c"],
            "col2": [1.0, 2.0, np.nan],
            "col3": [1.0, 2.0, np.nan],
        },
    )
    with pytest.raises(
        TypeError,
        match=(
            f"Passing 'result_names' as a {type(result_names)} is not "
            "supported. Provide 'result_names' as a tuple instead."
        ),
    ):
        df1.compare(df2, result_names=result_names)


@pytest.mark.parametrize(
    "val1,val2",
    [(4, pd.NA), (pd.NA, pd.NA), (pd.NA, 4)],
)
def test_compare_ea_and_np_dtype(val1, val2):
    # GH 48966
    arr = [4.0, val1]
    ser = pd.Series([1, val2], dtype="Int64")

    df1 = pd.DataFrame({"a": arr, "b": [1.0, 2]})
    df2 = pd.DataFrame({"a": ser, "b": [1.0, 2]})
    expected = pd.DataFrame(
        {
            ("a", "self"): arr,
            ("a", "other"): ser,
            ("b", "self"): np.nan,
            ("b", "other"): np.nan,
        }
    )
    if val1 is pd.NA and val2 is pd.NA:
        # GH#18463 TODO: is this really the desired behavior?
        expected.loc[1, ("a", "self")] = np.nan

    if val1 is pd.NA and np_version_gte1p25:
        # can't compare with numpy array if it contains pd.NA
        with pytest.raises(TypeError, match="boolean value of NA is ambiguous"):
            result = df1.compare(df2, keep_shape=True)
    else:
        result = df1.compare(df2, keep_shape=True)
        tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "df1_val,df2_val,diff_self,diff_other",
    [
        (4, 3, 4, 3),
        (4, 4, pd.NA, pd.NA),
        (4, pd.NA, 4, pd.NA),
        (pd.NA, pd.NA, pd.NA, pd.NA),
    ],
)
def test_compare_nullable_int64_dtype(df1_val, df2_val, diff_self, diff_other):
    # GH 48966
    df1 = pd.DataFrame({"a": pd.Series([df1_val, pd.NA], dtype="Int64"), "b": [1.0, 2]})
    df2 = df1.copy()
    df2.loc[0, "a"] = df2_val

    expected = pd.DataFrame(
        {
            ("a", "self"): pd.Series([diff_self, pd.NA], dtype="Int64"),
            ("a", "other"): pd.Series([diff_other, pd.NA], dtype="Int64"),
            ("b", "self"): np.nan,
            ("b", "other"): np.nan,
        }
    )
    result = df1.compare(df2, keep_shape=True)
    tm.assert_frame_equal(result, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\reshape\test_qcut.py ===
import os

import numpy as np
import pytest

import pandas as pd
from pandas import (
    Categorical,
    DatetimeIndex,
    Interval,
    IntervalIndex,
    NaT,
    Series,
    Timedelta,
    TimedeltaIndex,
    Timestamp,
    cut,
    date_range,
    isna,
    qcut,
    timedelta_range,
)
import pandas._testing as tm
from pandas.api.types import CategoricalDtype

from pandas.tseries.offsets import Day


def test_qcut():
    arr = np.random.default_rng(2).standard_normal(1000)

    # We store the bins as Index that have been
    # rounded to comparisons are a bit tricky.
    labels, _ = qcut(arr, 4, retbins=True)
    ex_bins = np.quantile(arr, [0, 0.25, 0.5, 0.75, 1.0])

    result = labels.categories.left.values
    assert np.allclose(result, ex_bins[:-1], atol=1e-2)

    result = labels.categories.right.values
    assert np.allclose(result, ex_bins[1:], atol=1e-2)

    ex_levels = cut(arr, ex_bins, include_lowest=True)
    tm.assert_categorical_equal(labels, ex_levels)


def test_qcut_bounds():
    arr = np.random.default_rng(2).standard_normal(1000)

    factor = qcut(arr, 10, labels=False)
    assert len(np.unique(factor)) == 10


def test_qcut_specify_quantiles():
    arr = np.random.default_rng(2).standard_normal(100)
    factor = qcut(arr, [0, 0.25, 0.5, 0.75, 1.0])

    expected = qcut(arr, 4)
    tm.assert_categorical_equal(factor, expected)


def test_qcut_all_bins_same():
    with pytest.raises(ValueError, match="edges.*unique"):
        qcut([0, 0, 0, 0, 0, 0, 0, 0, 0, 0], 3)


def test_qcut_include_lowest():
    values = np.arange(10)
    ii = qcut(values, 4)

    ex_levels = IntervalIndex(
        [
            Interval(-0.001, 2.25),
            Interval(2.25, 4.5),
            Interval(4.5, 6.75),
            Interval(6.75, 9),
        ]
    )
    tm.assert_index_equal(ii.categories, ex_levels)


def test_qcut_nas():
    arr = np.random.default_rng(2).standard_normal(100)
    arr[:20] = np.nan

    result = qcut(arr, 4)
    assert isna(result[:20]).all()


def test_qcut_index():
    result = qcut([0, 2], 2)
    intervals = [Interval(-0.001, 1), Interval(1, 2)]

    expected = Categorical(intervals, ordered=True)
    tm.assert_categorical_equal(result, expected)


def test_qcut_binning_issues(datapath):
    # see gh-1978, gh-1979
    cut_file = datapath(os.path.join("reshape", "data", "cut_data.csv"))
    arr = np.loadtxt(cut_file)
    result = qcut(arr, 20)

    starts = []
    ends = []

    for lev in np.unique(result):
        s = lev.left
        e = lev.right
        assert s != e

        starts.append(float(s))
        ends.append(float(e))

    for (sp, sn), (ep, en) in zip(
        zip(starts[:-1], starts[1:]), zip(ends[:-1], ends[1:])
    ):
        assert sp < sn
        assert ep < en
        assert ep <= sn


def test_qcut_return_intervals():
    ser = Series([0, 1, 2, 3, 4, 5, 6, 7, 8])
    res = qcut(ser, [0, 0.333, 0.666, 1])

    exp_levels = np.array(
        [Interval(-0.001, 2.664), Interval(2.664, 5.328), Interval(5.328, 8)]
    )
    exp = Series(exp_levels.take([0, 0, 0, 1, 1, 1, 2, 2, 2])).astype(
        CategoricalDtype(ordered=True)
    )
    tm.assert_series_equal(res, exp)


@pytest.mark.parametrize("labels", ["foo", 1, True])
def test_qcut_incorrect_labels(labels):
    # GH 13318
    values = range(5)
    msg = "Bin labels must either be False, None or passed in as a list-like argument"
    with pytest.raises(ValueError, match=msg):
        qcut(values, 4, labels=labels)


@pytest.mark.parametrize("labels", [["a", "b", "c"], list(range(3))])
def test_qcut_wrong_length_labels(labels):
    # GH 13318
    values = range(10)
    msg = "Bin labels must be one fewer than the number of bin edges"
    with pytest.raises(ValueError, match=msg):
        qcut(values, 4, labels=labels)


@pytest.mark.parametrize(
    "labels, expected",
    [
        (["a", "b", "c"], Categorical(["a", "b", "c"], ordered=True)),
        (list(range(3)), Categorical([0, 1, 2], ordered=True)),
    ],
)
def test_qcut_list_like_labels(labels, expected):
    # GH 13318
    values = range(3)
    result = qcut(values, 3, labels=labels)
    tm.assert_categorical_equal(result, expected)


@pytest.mark.parametrize(
    "kwargs,msg",
    [
        ({"duplicates": "drop"}, None),
        ({}, "Bin edges must be unique"),
        ({"duplicates": "raise"}, "Bin edges must be unique"),
        ({"duplicates": "foo"}, "invalid value for 'duplicates' parameter"),
    ],
)
def test_qcut_duplicates_bin(kwargs, msg):
    # see gh-7751
    values = [0, 0, 0, 0, 1, 2, 3]

    if msg is not None:
        with pytest.raises(ValueError, match=msg):
            qcut(values, 3, **kwargs)
    else:
        result = qcut(values, 3, **kwargs)
        expected = IntervalIndex([Interval(-0.001, 1), Interval(1, 3)])
        tm.assert_index_equal(result.categories, expected)


@pytest.mark.parametrize(
    "data,start,end", [(9.0, 8.999, 9.0), (0.0, -0.001, 0.0), (-9.0, -9.001, -9.0)]
)
@pytest.mark.parametrize("length", [1, 2])
@pytest.mark.parametrize("labels", [None, False])
def test_single_quantile(data, start, end, length, labels):
    # see gh-15431
    ser = Series([data] * length)
    result = qcut(ser, 1, labels=labels)

    if labels is None:
        intervals = IntervalIndex([Interval(start, end)] * length, closed="right")
        expected = Series(intervals).astype(CategoricalDtype(ordered=True))
    else:
        expected = Series([0] * length, dtype=np.intp)

    tm.assert_series_equal(result, expected)


@pytest.mark.parametrize(
    "ser",
    [
        Series(DatetimeIndex(["20180101", NaT, "20180103"])),
        Series(TimedeltaIndex(["0 days", NaT, "2 days"])),
    ],
    ids=lambda x: str(x.dtype),
)
def test_qcut_nat(ser, unit):
    # see gh-19768
    ser = ser.dt.as_unit(unit)
    td = Timedelta(1, unit=unit).as_unit(unit)

    left = Series([ser[0] - td, np.nan, ser[2] - Day()], dtype=ser.dtype)
    right = Series([ser[2] - Day(), np.nan, ser[2]], dtype=ser.dtype)
    intervals = IntervalIndex.from_arrays(left, right)
    expected = Series(Categorical(intervals, ordered=True))

    result = qcut(ser, 2)
    tm.assert_series_equal(result, expected)


@pytest.mark.parametrize("bins", [3, np.linspace(0, 1, 4)])
def test_datetime_tz_qcut(bins):
    # see gh-19872
    tz = "US/Eastern"
    ser = Series(date_range("20130101", periods=3, tz=tz))

    result = qcut(ser, bins)
    expected = Series(
        IntervalIndex(
            [
                Interval(
                    Timestamp("2012-12-31 23:59:59.999999999", tz=tz),
                    Timestamp("2013-01-01 16:00:00", tz=tz),
                ),
                Interval(
                    Timestamp("2013-01-01 16:00:00", tz=tz),
                    Timestamp("2013-01-02 08:00:00", tz=tz),
                ),
                Interval(
                    Timestamp("2013-01-02 08:00:00", tz=tz),
                    Timestamp("2013-01-03 00:00:00", tz=tz),
                ),
            ]
        )
    ).astype(CategoricalDtype(ordered=True))
    tm.assert_series_equal(result, expected)


@pytest.mark.parametrize(
    "arg,expected_bins",
    [
        [
            timedelta_range("1day", periods=3),
            TimedeltaIndex(["1 days", "2 days", "3 days"]),
        ],
        [
            date_range("20180101", periods=3),
            DatetimeIndex(["2018-01-01", "2018-01-02", "2018-01-03"]),
        ],
    ],
)
def test_date_like_qcut_bins(arg, expected_bins):
    # see gh-19891
    ser = Series(arg)
    result, result_bins = qcut(ser, 2, retbins=True)
    tm.assert_index_equal(result_bins, expected_bins)


@pytest.mark.parametrize("bins", [6, 7])
@pytest.mark.parametrize(
    "box, compare",
    [
        (Series, tm.assert_series_equal),
        (np.array, tm.assert_categorical_equal),
        (list, tm.assert_equal),
    ],
)
def test_qcut_bool_coercion_to_int(bins, box, compare):
    # issue 20303
    data_expected = box([0, 1, 1, 0, 1] * 10)
    data_result = box([False, True, True, False, True] * 10)
    expected = qcut(data_expected, bins, duplicates="drop")
    result = qcut(data_result, bins, duplicates="drop")
    compare(result, expected)


@pytest.mark.parametrize("q", [2, 5, 10])
def test_qcut_nullable_integer(q, any_numeric_ea_dtype):
    arr = pd.array(np.arange(100), dtype=any_numeric_ea_dtype)
    arr[::2] = pd.NA

    result = qcut(arr, q)
    expected = qcut(arr.astype(float), q)

    tm.assert_categorical_equal(result, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\io\parser\common\test_index.py ===
"""
Tests that work on both the Python and C engines but do not have a
specific classification into the other test modules.
"""
from datetime import datetime
from io import StringIO
import os

import pytest

from pandas import (
    DataFrame,
    Index,
    MultiIndex,
)
import pandas._testing as tm

pytestmark = pytest.mark.filterwarnings(
    "ignore:Passing a BlockManager to DataFrame:DeprecationWarning"
)

xfail_pyarrow = pytest.mark.usefixtures("pyarrow_xfail")
skip_pyarrow = pytest.mark.usefixtures("pyarrow_skip")


@pytest.mark.parametrize(
    "data,kwargs,expected",
    [
        (
            """foo,2,3,4,5
bar,7,8,9,10
baz,12,13,14,15
qux,12,13,14,15
foo2,12,13,14,15
bar2,12,13,14,15
""",
            {"index_col": 0, "names": ["index", "A", "B", "C", "D"]},
            DataFrame(
                [
                    [2, 3, 4, 5],
                    [7, 8, 9, 10],
                    [12, 13, 14, 15],
                    [12, 13, 14, 15],
                    [12, 13, 14, 15],
                    [12, 13, 14, 15],
                ],
                index=Index(["foo", "bar", "baz", "qux", "foo2", "bar2"], name="index"),
                columns=["A", "B", "C", "D"],
            ),
        ),
        (
            """foo,one,2,3,4,5
foo,two,7,8,9,10
foo,three,12,13,14,15
bar,one,12,13,14,15
bar,two,12,13,14,15
""",
            {"index_col": [0, 1], "names": ["index1", "index2", "A", "B", "C", "D"]},
            DataFrame(
                [
                    [2, 3, 4, 5],
                    [7, 8, 9, 10],
                    [12, 13, 14, 15],
                    [12, 13, 14, 15],
                    [12, 13, 14, 15],
                ],
                index=MultiIndex.from_tuples(
                    [
                        ("foo", "one"),
                        ("foo", "two"),
                        ("foo", "three"),
                        ("bar", "one"),
                        ("bar", "two"),
                    ],
                    names=["index1", "index2"],
                ),
                columns=["A", "B", "C", "D"],
            ),
        ),
    ],
)
def test_pass_names_with_index(all_parsers, data, kwargs, expected):
    parser = all_parsers
    result = parser.read_csv(StringIO(data), **kwargs)
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("index_col", [[0, 1], [1, 0]])
def test_multi_index_no_level_names(
    request, all_parsers, index_col, using_infer_string
):
    data = """index1,index2,A,B,C,D
foo,one,2,3,4,5
foo,two,7,8,9,10
foo,three,12,13,14,15
bar,one,12,13,14,15
bar,two,12,13,14,15
"""
    headless_data = "\n".join(data.split("\n")[1:])

    names = ["A", "B", "C", "D"]
    parser = all_parsers

    result = parser.read_csv(
        StringIO(headless_data), index_col=index_col, header=None, names=names
    )
    expected = parser.read_csv(StringIO(data), index_col=index_col)

    # No index names in headless data.
    expected.index.names = [None] * 2
    tm.assert_frame_equal(result, expected)


@skip_pyarrow
def test_multi_index_no_level_names_implicit(all_parsers):
    parser = all_parsers
    data = """A,B,C,D
foo,one,2,3,4,5
foo,two,7,8,9,10
foo,three,12,13,14,15
bar,one,12,13,14,15
bar,two,12,13,14,15
"""

    result = parser.read_csv(StringIO(data))
    expected = DataFrame(
        [
            [2, 3, 4, 5],
            [7, 8, 9, 10],
            [12, 13, 14, 15],
            [12, 13, 14, 15],
            [12, 13, 14, 15],
        ],
        columns=["A", "B", "C", "D"],
        index=MultiIndex.from_tuples(
            [
                ("foo", "one"),
                ("foo", "two"),
                ("foo", "three"),
                ("bar", "one"),
                ("bar", "two"),
            ]
        ),
    )
    tm.assert_frame_equal(result, expected)


@xfail_pyarrow  # TypeError: an integer is required
@pytest.mark.parametrize(
    "data,expected,header",
    [
        ("a,b", DataFrame(columns=["a", "b"]), [0]),
        (
            "a,b\nc,d",
            DataFrame(columns=MultiIndex.from_tuples([("a", "c"), ("b", "d")])),
            [0, 1],
        ),
    ],
)
@pytest.mark.parametrize("round_trip", [True, False])
def test_multi_index_blank_df(all_parsers, data, expected, header, round_trip):
    # see gh-14545
    parser = all_parsers
    data = expected.to_csv(index=False) if round_trip else data

    result = parser.read_csv(StringIO(data), header=header)
    tm.assert_frame_equal(result, expected)


@xfail_pyarrow  # AssertionError: DataFrame.columns are different
def test_no_unnamed_index(all_parsers):
    parser = all_parsers
    data = """ id c0 c1 c2
0 1 0 a b
1 2 0 c d
2 2 2 e f
"""
    result = parser.read_csv(StringIO(data), sep=" ")
    expected = DataFrame(
        [[0, 1, 0, "a", "b"], [1, 2, 0, "c", "d"], [2, 2, 2, "e", "f"]],
        columns=["Unnamed: 0", "id", "c0", "c1", "c2"],
    )
    tm.assert_frame_equal(result, expected)


def test_read_duplicate_index_explicit(all_parsers):
    data = """index,A,B,C,D
foo,2,3,4,5
bar,7,8,9,10
baz,12,13,14,15
qux,12,13,14,15
foo,12,13,14,15
bar,12,13,14,15
"""
    parser = all_parsers
    result = parser.read_csv(StringIO(data), index_col=0)

    expected = DataFrame(
        [
            [2, 3, 4, 5],
            [7, 8, 9, 10],
            [12, 13, 14, 15],
            [12, 13, 14, 15],
            [12, 13, 14, 15],
            [12, 13, 14, 15],
        ],
        columns=["A", "B", "C", "D"],
        index=Index(["foo", "bar", "baz", "qux", "foo", "bar"], name="index"),
    )
    tm.assert_frame_equal(result, expected)


@skip_pyarrow
def test_read_duplicate_index_implicit(all_parsers):
    data = """A,B,C,D
foo,2,3,4,5
bar,7,8,9,10
baz,12,13,14,15
qux,12,13,14,15
foo,12,13,14,15
bar,12,13,14,15
"""
    parser = all_parsers
    result = parser.read_csv(StringIO(data))

    expected = DataFrame(
        [
            [2, 3, 4, 5],
            [7, 8, 9, 10],
            [12, 13, 14, 15],
            [12, 13, 14, 15],
            [12, 13, 14, 15],
            [12, 13, 14, 15],
        ],
        columns=["A", "B", "C", "D"],
        index=Index(["foo", "bar", "baz", "qux", "foo", "bar"]),
    )
    tm.assert_frame_equal(result, expected)


@skip_pyarrow
def test_read_csv_no_index_name(all_parsers, csv_dir_path):
    parser = all_parsers
    csv2 = os.path.join(csv_dir_path, "test2.csv")
    result = parser.read_csv(csv2, index_col=0, parse_dates=True)

    expected = DataFrame(
        [
            [0.980269, 3.685731, -0.364216805298, -1.159738, "foo"],
            [1.047916, -0.041232, -0.16181208307, 0.212549, "bar"],
            [0.498581, 0.731168, -0.537677223318, 1.346270, "baz"],
            [1.120202, 1.567621, 0.00364077397681, 0.675253, "qux"],
            [-0.487094, 0.571455, -1.6116394093, 0.103469, "foo2"],
        ],
        columns=["A", "B", "C", "D", "E"],
        index=Index(
            [
                datetime(2000, 1, 3),
                datetime(2000, 1, 4),
                datetime(2000, 1, 5),
                datetime(2000, 1, 6),
                datetime(2000, 1, 7),
            ]
        ),
    )
    tm.assert_frame_equal(result, expected)


@skip_pyarrow
def test_empty_with_index(all_parsers):
    # see gh-10184
    data = "x,y"
    parser = all_parsers
    result = parser.read_csv(StringIO(data), index_col=0)

    expected = DataFrame(columns=["y"], index=Index([], name="x"))
    tm.assert_frame_equal(result, expected)


# CSV parse error: Empty CSV file or block: cannot infer number of columns
@skip_pyarrow
def test_empty_with_multi_index(all_parsers):
    # see gh-10467
    data = "x,y,z"
    parser = all_parsers
    result = parser.read_csv(StringIO(data), index_col=["x", "y"])

    expected = DataFrame(
        columns=["z"], index=MultiIndex.from_arrays([[]] * 2, names=["x", "y"])
    )
    tm.assert_frame_equal(result, expected)


# CSV parse error: Empty CSV file or block: cannot infer number of columns
@skip_pyarrow
def test_empty_with_reversed_multi_index(all_parsers):
    data = "x,y,z"
    parser = all_parsers
    result = parser.read_csv(StringIO(data), index_col=[1, 0])

    expected = DataFrame(
        columns=["z"], index=MultiIndex.from_arrays([[]] * 2, names=["y", "x"])
    )
    tm.assert_frame_equal(result, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\frame\methods\test_explode.py ===
import re

import numpy as np
import pytest

import pandas as pd
import pandas._testing as tm


def test_error():
    df = pd.DataFrame(
        {"A": pd.Series([[0, 1, 2], np.nan, [], (3, 4)], index=list("abcd")), "B": 1}
    )
    with pytest.raises(
        ValueError, match="column must be a scalar, tuple, or list thereof"
    ):
        df.explode([list("AA")])

    with pytest.raises(ValueError, match="column must be unique"):
        df.explode(list("AA"))

    df.columns = list("AA")
    with pytest.raises(
        ValueError,
        match=re.escape("DataFrame columns must be unique. Duplicate columns: ['A']"),
    ):
        df.explode("A")


@pytest.mark.parametrize(
    "input_subset, error_message",
    [
        (
            list("AC"),
            "columns must have matching element counts",
        ),
        (
            [],
            "column must be nonempty",
        ),
        (
            list("AC"),
            "columns must have matching element counts",
        ),
    ],
)
def test_error_multi_columns(input_subset, error_message):
    # GH 39240
    df = pd.DataFrame(
        {
            "A": [[0, 1, 2], np.nan, [], (3, 4)],
            "B": 1,
            "C": [["a", "b", "c"], "foo", [], ["d", "e", "f"]],
        },
        index=list("abcd"),
    )
    with pytest.raises(ValueError, match=error_message):
        df.explode(input_subset)


@pytest.mark.parametrize(
    "scalar",
    ["a", 0, 1.5, pd.Timedelta("1 days"), pd.Timestamp("2019-12-31")],
)
def test_basic(scalar):
    df = pd.DataFrame(
        {scalar: pd.Series([[0, 1, 2], np.nan, [], (3, 4)], index=list("abcd")), "B": 1}
    )
    result = df.explode(scalar)
    expected = pd.DataFrame(
        {
            scalar: pd.Series(
                [0, 1, 2, np.nan, np.nan, 3, 4], index=list("aaabcdd"), dtype=object
            ),
            "B": 1,
        }
    )
    tm.assert_frame_equal(result, expected)


def test_multi_index_rows():
    df = pd.DataFrame(
        {"A": np.array([[0, 1, 2], np.nan, [], (3, 4)], dtype=object), "B": 1},
        index=pd.MultiIndex.from_tuples([("a", 1), ("a", 2), ("b", 1), ("b", 2)]),
    )

    result = df.explode("A")
    expected = pd.DataFrame(
        {
            "A": pd.Series(
                [0, 1, 2, np.nan, np.nan, 3, 4],
                index=pd.MultiIndex.from_tuples(
                    [
                        ("a", 1),
                        ("a", 1),
                        ("a", 1),
                        ("a", 2),
                        ("b", 1),
                        ("b", 2),
                        ("b", 2),
                    ]
                ),
                dtype=object,
            ),
            "B": 1,
        }
    )
    tm.assert_frame_equal(result, expected)


def test_multi_index_columns():
    df = pd.DataFrame(
        {("A", 1): np.array([[0, 1, 2], np.nan, [], (3, 4)], dtype=object), ("A", 2): 1}
    )

    result = df.explode(("A", 1))
    expected = pd.DataFrame(
        {
            ("A", 1): pd.Series(
                [0, 1, 2, np.nan, np.nan, 3, 4],
                index=pd.Index([0, 0, 0, 1, 2, 3, 3]),
                dtype=object,
            ),
            ("A", 2): 1,
        }
    )
    tm.assert_frame_equal(result, expected)


def test_usecase():
    # explode a single column
    # gh-10511
    df = pd.DataFrame(
        [[11, range(5), 10], [22, range(3), 20]], columns=list("ABC")
    ).set_index("C")
    result = df.explode("B")

    expected = pd.DataFrame(
        {
            "A": [11, 11, 11, 11, 11, 22, 22, 22],
            "B": np.array([0, 1, 2, 3, 4, 0, 1, 2], dtype=object),
            "C": [10, 10, 10, 10, 10, 20, 20, 20],
        },
        columns=list("ABC"),
    ).set_index("C")

    tm.assert_frame_equal(result, expected)

    # gh-8517
    df = pd.DataFrame(
        [["2014-01-01", "Alice", "A B"], ["2014-01-02", "Bob", "C D"]],
        columns=["dt", "name", "text"],
    )
    result = df.assign(text=df.text.str.split(" ")).explode("text")
    expected = pd.DataFrame(
        [
            ["2014-01-01", "Alice", "A"],
            ["2014-01-01", "Alice", "B"],
            ["2014-01-02", "Bob", "C"],
            ["2014-01-02", "Bob", "D"],
        ],
        columns=["dt", "name", "text"],
        index=[0, 0, 1, 1],
    )
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "input_dict, input_index, expected_dict, expected_index",
    [
        (
            {"col1": [[1, 2], [3, 4]], "col2": ["foo", "bar"]},
            [0, 0],
            {"col1": [1, 2, 3, 4], "col2": ["foo", "foo", "bar", "bar"]},
            [0, 0, 0, 0],
        ),
        (
            {"col1": [[1, 2], [3, 4]], "col2": ["foo", "bar"]},
            pd.Index([0, 0], name="my_index"),
            {"col1": [1, 2, 3, 4], "col2": ["foo", "foo", "bar", "bar"]},
            pd.Index([0, 0, 0, 0], name="my_index"),
        ),
        (
            {"col1": [[1, 2], [3, 4]], "col2": ["foo", "bar"]},
            pd.MultiIndex.from_arrays(
                [[0, 0], [1, 1]], names=["my_first_index", "my_second_index"]
            ),
            {"col1": [1, 2, 3, 4], "col2": ["foo", "foo", "bar", "bar"]},
            pd.MultiIndex.from_arrays(
                [[0, 0, 0, 0], [1, 1, 1, 1]],
                names=["my_first_index", "my_second_index"],
            ),
        ),
        (
            {"col1": [[1, 2], [3, 4]], "col2": ["foo", "bar"]},
            pd.MultiIndex.from_arrays([[0, 0], [1, 1]], names=["my_index", None]),
            {"col1": [1, 2, 3, 4], "col2": ["foo", "foo", "bar", "bar"]},
            pd.MultiIndex.from_arrays(
                [[0, 0, 0, 0], [1, 1, 1, 1]], names=["my_index", None]
            ),
        ),
    ],
)
def test_duplicate_index(input_dict, input_index, expected_dict, expected_index):
    # GH 28005
    df = pd.DataFrame(input_dict, index=input_index, dtype=object)
    result = df.explode("col1")
    expected = pd.DataFrame(expected_dict, index=expected_index, dtype=object)
    tm.assert_frame_equal(result, expected)


def test_ignore_index():
    # GH 34932
    df = pd.DataFrame({"id": range(0, 20, 10), "values": [list("ab"), list("cd")]})
    result = df.explode("values", ignore_index=True)
    expected = pd.DataFrame(
        {"id": [0, 0, 10, 10], "values": list("abcd")}, index=[0, 1, 2, 3]
    )
    tm.assert_frame_equal(result, expected)


def test_explode_sets():
    # https://github.com/pandas-dev/pandas/issues/35614
    df = pd.DataFrame({"a": [{"x", "y"}], "b": [1]}, index=[1])
    result = df.explode(column="a").sort_values(by="a")
    expected = pd.DataFrame({"a": ["x", "y"], "b": [1, 1]}, index=[1, 1])
    tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize(
    "input_subset, expected_dict, expected_index",
    [
        (
            list("AC"),
            {
                "A": pd.Series(
                    [0, 1, 2, np.nan, np.nan, 3, 4, np.nan],
                    index=list("aaabcdde"),
                    dtype=object,
                ),
                "B": 1,
                "C": ["a", "b", "c", "foo", np.nan, "d", "e", np.nan],
            },
            list("aaabcdde"),
        ),
        (
            list("A"),
            {
                "A": pd.Series(
                    [0, 1, 2, np.nan, np.nan, 3, 4, np.nan],
                    index=list("aaabcdde"),
                    dtype=object,
                ),
                "B": 1,
                "C": [
                    ["a", "b", "c"],
                    ["a", "b", "c"],
                    ["a", "b", "c"],
                    "foo",
                    [],
                    ["d", "e"],
                    ["d", "e"],
                    np.nan,
                ],
            },
            list("aaabcdde"),
        ),
    ],
)
def test_multi_columns(input_subset, expected_dict, expected_index):
    # GH 39240
    df = pd.DataFrame(
        {
            "A": [[0, 1, 2], np.nan, [], (3, 4), np.nan],
            "B": 1,
            "C": [["a", "b", "c"], "foo", [], ["d", "e"], np.nan],
        },
        index=list("abcde"),
    )
    result = df.explode(input_subset)
    expected = pd.DataFrame(expected_dict, expected_index)
    tm.assert_frame_equal(result, expected)


def test_multi_columns_nan_empty():
    # GH 46084
    df = pd.DataFrame(
        {
            "A": [[0, 1], [5], [], [2, 3]],
            "B": [9, 8, 7, 6],
            "C": [[1, 2], np.nan, [], [3, 4]],
        }
    )
    result = df.explode(["A", "C"])
    expected = pd.DataFrame(
        {
            "A": np.array([0, 1, 5, np.nan, 2, 3], dtype=object),
            "B": [9, 9, 8, 7, 6, 6],
            "C": np.array([1, 2, np.nan, np.nan, 3, 4], dtype=object),
        },
        index=[0, 0, 1, 2, 3, 3],
    )
    tm.assert_frame_equal(result, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\indexing\test_scalar.py ===
""" test scalar indexing, including at and iat """
from datetime import (
    datetime,
    timedelta,
)
import itertools

import numpy as np
import pytest

from pandas import (
    DataFrame,
    Series,
    Timedelta,
    Timestamp,
    date_range,
)
import pandas._testing as tm


def generate_indices(f, values=False):
    """
    generate the indices
    if values is True , use the axis values
    is False, use the range
    """
    axes = f.axes
    if values:
        axes = (list(range(len(ax))) for ax in axes)

    return itertools.product(*axes)


class TestScalar:
    @pytest.mark.parametrize("kind", ["series", "frame"])
    @pytest.mark.parametrize("col", ["ints", "uints"])
    def test_iat_set_ints(self, kind, col, request):
        f = request.getfixturevalue(f"{kind}_{col}")
        indices = generate_indices(f, True)
        for i in indices:
            f.iat[i] = 1
            expected = f.values[i]
            tm.assert_almost_equal(expected, 1)

    @pytest.mark.parametrize("kind", ["series", "frame"])
    @pytest.mark.parametrize("col", ["labels", "ts", "floats"])
    def test_iat_set_other(self, kind, col, request):
        f = request.getfixturevalue(f"{kind}_{col}")
        msg = "iAt based indexing can only have integer indexers"
        with pytest.raises(ValueError, match=msg):
            idx = next(generate_indices(f, False))
            f.iat[idx] = 1

    @pytest.mark.parametrize("kind", ["series", "frame"])
    @pytest.mark.parametrize("col", ["ints", "uints", "labels", "ts", "floats"])
    def test_at_set_ints_other(self, kind, col, request):
        f = request.getfixturevalue(f"{kind}_{col}")
        indices = generate_indices(f, False)
        for i in indices:
            f.at[i] = 1
            expected = f.loc[i]
            tm.assert_almost_equal(expected, 1)


class TestAtAndiAT:
    # at and iat tests that don't need Base class

    def test_float_index_at_iat(self):
        ser = Series([1, 2, 3], index=[0.1, 0.2, 0.3])
        for el, item in ser.items():
            assert ser.at[el] == item
        for i in range(len(ser)):
            assert ser.iat[i] == i + 1

    def test_at_iat_coercion(self):
        # as timestamp is not a tuple!
        dates = date_range("1/1/2000", periods=8)
        df = DataFrame(
            np.random.default_rng(2).standard_normal((8, 4)),
            index=dates,
            columns=["A", "B", "C", "D"],
        )
        s = df["A"]

        result = s.at[dates[5]]
        xp = s.values[5]
        assert result == xp

    @pytest.mark.parametrize(
        "ser, expected",
        [
            [
                Series(["2014-01-01", "2014-02-02"], dtype="datetime64[ns]"),
                Timestamp("2014-02-02"),
            ],
            [
                Series(["1 days", "2 days"], dtype="timedelta64[ns]"),
                Timedelta("2 days"),
            ],
        ],
    )
    def test_iloc_iat_coercion_datelike(self, indexer_ial, ser, expected):
        # GH 7729
        # make sure we are boxing the returns
        result = indexer_ial(ser)[1]
        assert result == expected

    def test_imethods_with_dups(self):
        # GH6493
        # iat/iloc with dups

        s = Series(range(5), index=[1, 1, 2, 2, 3], dtype="int64")
        result = s.iloc[2]
        assert result == 2
        result = s.iat[2]
        assert result == 2

        msg = "index 10 is out of bounds for axis 0 with size 5"
        with pytest.raises(IndexError, match=msg):
            s.iat[10]
        msg = "index -10 is out of bounds for axis 0 with size 5"
        with pytest.raises(IndexError, match=msg):
            s.iat[-10]

        result = s.iloc[[2, 3]]
        expected = Series([2, 3], [2, 2], dtype="int64")
        tm.assert_series_equal(result, expected)

        df = s.to_frame()
        result = df.iloc[2]
        expected = Series(2, index=[0], name=2)
        tm.assert_series_equal(result, expected)

        result = df.iat[2, 0]
        assert result == 2

    def test_frame_at_with_duplicate_axes(self):
        # GH#33041
        arr = np.random.default_rng(2).standard_normal(6).reshape(3, 2)
        df = DataFrame(arr, columns=["A", "A"])

        result = df.at[0, "A"]
        expected = df.iloc[0].copy()

        tm.assert_series_equal(result, expected)

        result = df.T.at["A", 0]
        tm.assert_series_equal(result, expected)

        # setter
        df.at[1, "A"] = 2
        expected = Series([2.0, 2.0], index=["A", "A"], name=1)
        tm.assert_series_equal(df.iloc[1], expected)

    def test_at_getitem_dt64tz_values(self):
        # gh-15822
        df = DataFrame(
            {
                "name": ["John", "Anderson"],
                "date": [
                    Timestamp(2017, 3, 13, 13, 32, 56),
                    Timestamp(2017, 2, 16, 12, 10, 3),
                ],
            }
        )
        df["date"] = df["date"].dt.tz_localize("Asia/Shanghai")

        expected = Timestamp("2017-03-13 13:32:56+0800", tz="Asia/Shanghai")

        result = df.loc[0, "date"]
        assert result == expected

        result = df.at[0, "date"]
        assert result == expected

    def test_mixed_index_at_iat_loc_iloc_series(self):
        # GH 19860
        s = Series([1, 2, 3, 4, 5], index=["a", "b", "c", 1, 2])
        for el, item in s.items():
            assert s.at[el] == s.loc[el] == item
        for i in range(len(s)):
            assert s.iat[i] == s.iloc[i] == i + 1

        with pytest.raises(KeyError, match="^4$"):
            s.at[4]
        with pytest.raises(KeyError, match="^4$"):
            s.loc[4]

    def test_mixed_index_at_iat_loc_iloc_dataframe(self):
        # GH 19860
        df = DataFrame(
            [[0, 1, 2, 3, 4], [5, 6, 7, 8, 9]], columns=["a", "b", "c", 1, 2]
        )
        for rowIdx, row in df.iterrows():
            for el, item in row.items():
                assert df.at[rowIdx, el] == df.loc[rowIdx, el] == item

        for row in range(2):
            for i in range(5):
                assert df.iat[row, i] == df.iloc[row, i] == row * 5 + i

        with pytest.raises(KeyError, match="^3$"):
            df.at[0, 3]
        with pytest.raises(KeyError, match="^3$"):
            df.loc[0, 3]

    def test_iat_setter_incompatible_assignment(self):
        # GH 23236
        result = DataFrame({"a": [0.0, 1.0], "b": [4, 5]})
        result.iat[0, 0] = None
        expected = DataFrame({"a": [None, 1], "b": [4, 5]})
        tm.assert_frame_equal(result, expected)


def test_iat_dont_wrap_object_datetimelike():
    # GH#32809 .iat calls go through DataFrame._get_value, should not
    #  call maybe_box_datetimelike
    dti = date_range("2016-01-01", periods=3)
    tdi = dti - dti
    ser = Series(dti.to_pydatetime(), dtype=object)
    ser2 = Series(tdi.to_pytimedelta(), dtype=object)
    df = DataFrame({"A": ser, "B": ser2})
    assert (df.dtypes == object).all()

    for result in [df.at[0, "A"], df.iat[0, 0], df.loc[0, "A"], df.iloc[0, 0]]:
        assert result is ser[0]
        assert isinstance(result, datetime)
        assert not isinstance(result, Timestamp)

    for result in [df.at[1, "B"], df.iat[1, 1], df.loc[1, "B"], df.iloc[1, 1]]:
        assert result is ser2[1]
        assert isinstance(result, timedelta)
        assert not isinstance(result, Timedelta)


def test_at_with_tuple_index_get():
    # GH 26989
    # DataFrame.at getter works with Index of tuples
    df = DataFrame({"a": [1, 2]}, index=[(1, 2), (3, 4)])
    assert df.index.nlevels == 1
    assert df.at[(1, 2), "a"] == 1

    # Series.at getter works with Index of tuples
    series = df["a"]
    assert series.index.nlevels == 1
    assert series.at[(1, 2)] == 1


@pytest.mark.filterwarnings("ignore:Setting a value on a view:FutureWarning")
def test_at_with_tuple_index_set():
    # GH 26989
    # DataFrame.at setter works with Index of tuples
    df = DataFrame({"a": [1, 2]}, index=[(1, 2), (3, 4)])
    assert df.index.nlevels == 1
    df.at[(1, 2), "a"] = 2
    assert df.at[(1, 2), "a"] == 2

    # Series.at setter works with Index of tuples
    series = df["a"]
    assert series.index.nlevels == 1
    series.at[1, 2] = 3
    assert series.at[1, 2] == 3


class TestMultiIndexScalar:
    def test_multiindex_at_get(self):
        # GH 26989
        # DataFrame.at and DataFrame.loc getter works with MultiIndex
        df = DataFrame({"a": [1, 2]}, index=[[1, 2], [3, 4]])
        assert df.index.nlevels == 2
        assert df.at[(1, 3), "a"] == 1
        assert df.loc[(1, 3), "a"] == 1

        # Series.at and Series.loc getter works with MultiIndex
        series = df["a"]
        assert series.index.nlevels == 2
        assert series.at[1, 3] == 1
        assert series.loc[1, 3] == 1

    @pytest.mark.filterwarnings("ignore:Setting a value on a view:FutureWarning")
    def test_multiindex_at_set(self):
        # GH 26989
        # DataFrame.at and DataFrame.loc setter works with MultiIndex
        df = DataFrame({"a": [1, 2]}, index=[[1, 2], [3, 4]])
        assert df.index.nlevels == 2
        df.at[(1, 3), "a"] = 3
        assert df.at[(1, 3), "a"] == 3
        df.loc[(1, 3), "a"] = 4
        assert df.loc[(1, 3), "a"] == 4

        # Series.at and Series.loc setter works with MultiIndex
        series = df["a"]
        assert series.index.nlevels == 2
        series.at[1, 3] = 5
        assert series.at[1, 3] == 5
        series.loc[1, 3] = 6
        assert series.loc[1, 3] == 6

    def test_multiindex_at_get_one_level(self):
        # GH#38053
        s2 = Series((0, 1), index=[[False, True]])
        result = s2.at[False]
        assert result == 0

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\tseries\offsets\test_quarter.py ===
"""
Tests for the following offsets:
- QuarterBegin
- QuarterEnd
"""
from __future__ import annotations

from datetime import datetime

import pytest

import pandas._testing as tm
from pandas.tests.tseries.offsets.common import (
    assert_is_on_offset,
    assert_offset_equal,
)

from pandas.tseries.offsets import (
    QuarterBegin,
    QuarterEnd,
)


@pytest.mark.parametrize("klass", (QuarterBegin, QuarterEnd))
def test_quarterly_dont_normalize(klass):
    date = datetime(2012, 3, 31, 5, 30)
    result = date + klass()
    assert result.time() == date.time()


@pytest.mark.parametrize("offset", [QuarterBegin(), QuarterEnd()])
@pytest.mark.parametrize(
    "date",
    [
        datetime(2016, m, d)
        for m in [10, 11, 12]
        for d in [1, 2, 3, 28, 29, 30, 31]
        if not (m == 11 and d == 31)
    ],
)
def test_on_offset(offset, date):
    res = offset.is_on_offset(date)
    slow_version = date == (date + offset) - offset
    assert res == slow_version


class TestQuarterBegin:
    def test_repr(self):
        expected = "<QuarterBegin: startingMonth=3>"
        assert repr(QuarterBegin()) == expected
        expected = "<QuarterBegin: startingMonth=3>"
        assert repr(QuarterBegin(startingMonth=3)) == expected
        expected = "<QuarterBegin: startingMonth=1>"
        assert repr(QuarterBegin(startingMonth=1)) == expected

    def test_is_anchored(self):
        msg = "QuarterBegin.is_anchored is deprecated "

        with tm.assert_produces_warning(FutureWarning, match=msg):
            assert QuarterBegin(startingMonth=1).is_anchored()
            assert QuarterBegin().is_anchored()
            assert not QuarterBegin(2, startingMonth=1).is_anchored()

    def test_offset_corner_case(self):
        # corner
        offset = QuarterBegin(n=-1, startingMonth=1)
        assert datetime(2010, 2, 1) + offset == datetime(2010, 1, 1)

    offset_cases = []
    offset_cases.append(
        (
            QuarterBegin(startingMonth=1),
            {
                datetime(2007, 12, 1): datetime(2008, 1, 1),
                datetime(2008, 1, 1): datetime(2008, 4, 1),
                datetime(2008, 2, 15): datetime(2008, 4, 1),
                datetime(2008, 2, 29): datetime(2008, 4, 1),
                datetime(2008, 3, 15): datetime(2008, 4, 1),
                datetime(2008, 3, 31): datetime(2008, 4, 1),
                datetime(2008, 4, 15): datetime(2008, 7, 1),
                datetime(2008, 4, 1): datetime(2008, 7, 1),
            },
        )
    )

    offset_cases.append(
        (
            QuarterBegin(startingMonth=2),
            {
                datetime(2008, 1, 1): datetime(2008, 2, 1),
                datetime(2008, 1, 31): datetime(2008, 2, 1),
                datetime(2008, 1, 15): datetime(2008, 2, 1),
                datetime(2008, 2, 29): datetime(2008, 5, 1),
                datetime(2008, 3, 15): datetime(2008, 5, 1),
                datetime(2008, 3, 31): datetime(2008, 5, 1),
                datetime(2008, 4, 15): datetime(2008, 5, 1),
                datetime(2008, 4, 30): datetime(2008, 5, 1),
            },
        )
    )

    offset_cases.append(
        (
            QuarterBegin(startingMonth=1, n=0),
            {
                datetime(2008, 1, 1): datetime(2008, 1, 1),
                datetime(2008, 12, 1): datetime(2009, 1, 1),
                datetime(2008, 1, 1): datetime(2008, 1, 1),
                datetime(2008, 2, 15): datetime(2008, 4, 1),
                datetime(2008, 2, 29): datetime(2008, 4, 1),
                datetime(2008, 3, 15): datetime(2008, 4, 1),
                datetime(2008, 3, 31): datetime(2008, 4, 1),
                datetime(2008, 4, 15): datetime(2008, 7, 1),
                datetime(2008, 4, 30): datetime(2008, 7, 1),
            },
        )
    )

    offset_cases.append(
        (
            QuarterBegin(startingMonth=1, n=-1),
            {
                datetime(2008, 1, 1): datetime(2007, 10, 1),
                datetime(2008, 1, 31): datetime(2008, 1, 1),
                datetime(2008, 2, 15): datetime(2008, 1, 1),
                datetime(2008, 2, 29): datetime(2008, 1, 1),
                datetime(2008, 3, 15): datetime(2008, 1, 1),
                datetime(2008, 3, 31): datetime(2008, 1, 1),
                datetime(2008, 4, 15): datetime(2008, 4, 1),
                datetime(2008, 4, 30): datetime(2008, 4, 1),
                datetime(2008, 7, 1): datetime(2008, 4, 1),
            },
        )
    )

    offset_cases.append(
        (
            QuarterBegin(startingMonth=1, n=2),
            {
                datetime(2008, 1, 1): datetime(2008, 7, 1),
                datetime(2008, 2, 15): datetime(2008, 7, 1),
                datetime(2008, 2, 29): datetime(2008, 7, 1),
                datetime(2008, 3, 15): datetime(2008, 7, 1),
                datetime(2008, 3, 31): datetime(2008, 7, 1),
                datetime(2008, 4, 15): datetime(2008, 10, 1),
                datetime(2008, 4, 1): datetime(2008, 10, 1),
            },
        )
    )

    @pytest.mark.parametrize("case", offset_cases)
    def test_offset(self, case):
        offset, cases = case
        for base, expected in cases.items():
            assert_offset_equal(offset, base, expected)


class TestQuarterEnd:
    def test_repr(self):
        expected = "<QuarterEnd: startingMonth=3>"
        assert repr(QuarterEnd()) == expected
        expected = "<QuarterEnd: startingMonth=3>"
        assert repr(QuarterEnd(startingMonth=3)) == expected
        expected = "<QuarterEnd: startingMonth=1>"
        assert repr(QuarterEnd(startingMonth=1)) == expected

    def test_is_anchored(self):
        msg = "QuarterEnd.is_anchored is deprecated "

        with tm.assert_produces_warning(FutureWarning, match=msg):
            assert QuarterEnd(startingMonth=1).is_anchored()
            assert QuarterEnd().is_anchored()
            assert not QuarterEnd(2, startingMonth=1).is_anchored()

    def test_offset_corner_case(self):
        # corner
        offset = QuarterEnd(n=-1, startingMonth=1)
        assert datetime(2010, 2, 1) + offset == datetime(2010, 1, 31)

    offset_cases = []
    offset_cases.append(
        (
            QuarterEnd(startingMonth=1),
            {
                datetime(2008, 1, 1): datetime(2008, 1, 31),
                datetime(2008, 1, 31): datetime(2008, 4, 30),
                datetime(2008, 2, 15): datetime(2008, 4, 30),
                datetime(2008, 2, 29): datetime(2008, 4, 30),
                datetime(2008, 3, 15): datetime(2008, 4, 30),
                datetime(2008, 3, 31): datetime(2008, 4, 30),
                datetime(2008, 4, 15): datetime(2008, 4, 30),
                datetime(2008, 4, 30): datetime(2008, 7, 31),
            },
        )
    )

    offset_cases.append(
        (
            QuarterEnd(startingMonth=2),
            {
                datetime(2008, 1, 1): datetime(2008, 2, 29),
                datetime(2008, 1, 31): datetime(2008, 2, 29),
                datetime(2008, 2, 15): datetime(2008, 2, 29),
                datetime(2008, 2, 29): datetime(2008, 5, 31),
                datetime(2008, 3, 15): datetime(2008, 5, 31),
                datetime(2008, 3, 31): datetime(2008, 5, 31),
                datetime(2008, 4, 15): datetime(2008, 5, 31),
                datetime(2008, 4, 30): datetime(2008, 5, 31),
            },
        )
    )

    offset_cases.append(
        (
            QuarterEnd(startingMonth=1, n=0),
            {
                datetime(2008, 1, 1): datetime(2008, 1, 31),
                datetime(2008, 1, 31): datetime(2008, 1, 31),
                datetime(2008, 2, 15): datetime(2008, 4, 30),
                datetime(2008, 2, 29): datetime(2008, 4, 30),
                datetime(2008, 3, 15): datetime(2008, 4, 30),
                datetime(2008, 3, 31): datetime(2008, 4, 30),
                datetime(2008, 4, 15): datetime(2008, 4, 30),
                datetime(2008, 4, 30): datetime(2008, 4, 30),
            },
        )
    )

    offset_cases.append(
        (
            QuarterEnd(startingMonth=1, n=-1),
            {
                datetime(2008, 1, 1): datetime(2007, 10, 31),
                datetime(2008, 1, 31): datetime(2007, 10, 31),
                datetime(2008, 2, 15): datetime(2008, 1, 31),
                datetime(2008, 2, 29): datetime(2008, 1, 31),
                datetime(2008, 3, 15): datetime(2008, 1, 31),
                datetime(2008, 3, 31): datetime(2008, 1, 31),
                datetime(2008, 4, 15): datetime(2008, 1, 31),
                datetime(2008, 4, 30): datetime(2008, 1, 31),
                datetime(2008, 7, 1): datetime(2008, 4, 30),
            },
        )
    )

    offset_cases.append(
        (
            QuarterEnd(startingMonth=1, n=2),
            {
                datetime(2008, 1, 31): datetime(2008, 7, 31),
                datetime(2008, 2, 15): datetime(2008, 7, 31),
                datetime(2008, 2, 29): datetime(2008, 7, 31),
                datetime(2008, 3, 15): datetime(2008, 7, 31),
                datetime(2008, 3, 31): datetime(2008, 7, 31),
                datetime(2008, 4, 15): datetime(2008, 7, 31),
                datetime(2008, 4, 30): datetime(2008, 10, 31),
            },
        )
    )

    @pytest.mark.parametrize("case", offset_cases)
    def test_offset(self, case):
        offset, cases = case
        for base, expected in cases.items():
            assert_offset_equal(offset, base, expected)

    on_offset_cases = [
        (QuarterEnd(1, startingMonth=1), datetime(2008, 1, 31), True),
        (QuarterEnd(1, startingMonth=1), datetime(2007, 12, 31), False),
        (QuarterEnd(1, startingMonth=1), datetime(2008, 2, 29), False),
        (QuarterEnd(1, startingMonth=1), datetime(2007, 3, 30), False),
        (QuarterEnd(1, startingMonth=1), datetime(2007, 3, 31), False),
        (QuarterEnd(1, startingMonth=1), datetime(2008, 4, 30), True),
        (QuarterEnd(1, startingMonth=1), datetime(2008, 5, 30), False),
        (QuarterEnd(1, startingMonth=1), datetime(2008, 5, 31), False),
        (QuarterEnd(1, startingMonth=1), datetime(2007, 6, 29), False),
        (QuarterEnd(1, startingMonth=1), datetime(2007, 6, 30), False),
        (QuarterEnd(1, startingMonth=2), datetime(2008, 1, 31), False),
        (QuarterEnd(1, startingMonth=2), datetime(2007, 12, 31), False),
        (QuarterEnd(1, startingMonth=2), datetime(2008, 2, 29), True),
        (QuarterEnd(1, startingMonth=2), datetime(2007, 3, 30), False),
        (QuarterEnd(1, startingMonth=2), datetime(2007, 3, 31), False),
        (QuarterEnd(1, startingMonth=2), datetime(2008, 4, 30), False),
        (QuarterEnd(1, startingMonth=2), datetime(2008, 5, 30), False),
        (QuarterEnd(1, startingMonth=2), datetime(2008, 5, 31), True),
        (QuarterEnd(1, startingMonth=2), datetime(2007, 6, 29), False),
        (QuarterEnd(1, startingMonth=2), datetime(2007, 6, 30), False),
        (QuarterEnd(1, startingMonth=3), datetime(2008, 1, 31), False),
        (QuarterEnd(1, startingMonth=3), datetime(2007, 12, 31), True),
        (QuarterEnd(1, startingMonth=3), datetime(2008, 2, 29), False),
        (QuarterEnd(1, startingMonth=3), datetime(2007, 3, 30), False),
        (QuarterEnd(1, startingMonth=3), datetime(2007, 3, 31), True),
        (QuarterEnd(1, startingMonth=3), datetime(2008, 4, 30), False),
        (QuarterEnd(1, startingMonth=3), datetime(2008, 5, 30), False),
        (QuarterEnd(1, startingMonth=3), datetime(2008, 5, 31), False),
        (QuarterEnd(1, startingMonth=3), datetime(2007, 6, 29), False),
        (QuarterEnd(1, startingMonth=3), datetime(2007, 6, 30), True),
    ]

    @pytest.mark.parametrize("case", on_offset_cases)
    def test_is_on_offset(self, case):
        offset, dt, expected = case
        assert_is_on_offset(offset, dt, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\arrays\sparse\test_indexing.py ===
import numpy as np
import pytest

import pandas as pd
from pandas import SparseDtype
import pandas._testing as tm
from pandas.core.arrays.sparse import SparseArray


@pytest.fixture
def arr_data():
    return np.array([np.nan, np.nan, 1, 2, 3, np.nan, 4, 5, np.nan, 6])


@pytest.fixture
def arr(arr_data):
    return SparseArray(arr_data)


class TestGetitem:
    def test_getitem(self, arr):
        dense = arr.to_dense()
        for i, value in enumerate(arr):
            tm.assert_almost_equal(value, dense[i])
            tm.assert_almost_equal(arr[-i], dense[-i])

    def test_getitem_arraylike_mask(self, arr):
        arr = SparseArray([0, 1, 2])
        result = arr[[True, False, True]]
        expected = SparseArray([0, 2])
        tm.assert_sp_array_equal(result, expected)

    @pytest.mark.parametrize(
        "slc",
        [
            np.s_[:],
            np.s_[1:10],
            np.s_[1:100],
            np.s_[10:1],
            np.s_[:-3],
            np.s_[-5:-4],
            np.s_[:-12],
            np.s_[-12:],
            np.s_[2:],
            np.s_[2::3],
            np.s_[::2],
            np.s_[::-1],
            np.s_[::-2],
            np.s_[1:6:2],
            np.s_[:-6:-2],
        ],
    )
    @pytest.mark.parametrize(
        "as_dense", [[np.nan] * 10, [1] * 10, [np.nan] * 5 + [1] * 5, []]
    )
    def test_getslice(self, slc, as_dense):
        as_dense = np.array(as_dense)
        arr = SparseArray(as_dense)

        result = arr[slc]
        expected = SparseArray(as_dense[slc])

        tm.assert_sp_array_equal(result, expected)

    def test_getslice_tuple(self):
        dense = np.array([np.nan, 0, 3, 4, 0, 5, np.nan, np.nan, 0])

        sparse = SparseArray(dense)
        res = sparse[(slice(4, None),)]
        exp = SparseArray(dense[4:])
        tm.assert_sp_array_equal(res, exp)

        sparse = SparseArray(dense, fill_value=0)
        res = sparse[(slice(4, None),)]
        exp = SparseArray(dense[4:], fill_value=0)
        tm.assert_sp_array_equal(res, exp)

        msg = "too many indices for array"
        with pytest.raises(IndexError, match=msg):
            sparse[4:, :]

        with pytest.raises(IndexError, match=msg):
            # check numpy compat
            dense[4:, :]

    def test_boolean_slice_empty(self):
        arr = SparseArray([0, 1, 2])
        res = arr[[False, False, False]]
        assert res.dtype == arr.dtype

    def test_getitem_bool_sparse_array(self, arr):
        # GH 23122
        spar_bool = SparseArray([False, True] * 5, dtype=np.bool_, fill_value=True)
        exp = SparseArray([np.nan, 2, np.nan, 5, 6])
        tm.assert_sp_array_equal(arr[spar_bool], exp)

        spar_bool = ~spar_bool
        res = arr[spar_bool]
        exp = SparseArray([np.nan, 1, 3, 4, np.nan])
        tm.assert_sp_array_equal(res, exp)

        spar_bool = SparseArray(
            [False, True, np.nan] * 3, dtype=np.bool_, fill_value=np.nan
        )
        res = arr[spar_bool]
        exp = SparseArray([np.nan, 3, 5])
        tm.assert_sp_array_equal(res, exp)

    def test_getitem_bool_sparse_array_as_comparison(self):
        # GH 45110
        arr = SparseArray([1, 2, 3, 4, np.nan, np.nan], fill_value=np.nan)
        res = arr[arr > 2]
        exp = SparseArray([3.0, 4.0], fill_value=np.nan)
        tm.assert_sp_array_equal(res, exp)

    def test_get_item(self, arr):
        zarr = SparseArray([0, 0, 1, 2, 3, 0, 4, 5, 0, 6], fill_value=0)

        assert np.isnan(arr[1])
        assert arr[2] == 1
        assert arr[7] == 5

        assert zarr[0] == 0
        assert zarr[2] == 1
        assert zarr[7] == 5

        errmsg = "must be an integer between -10 and 10"

        with pytest.raises(IndexError, match=errmsg):
            arr[11]

        with pytest.raises(IndexError, match=errmsg):
            arr[-11]

        assert arr[-1] == arr[len(arr) - 1]


class TestSetitem:
    def test_set_item(self, arr_data):
        arr = SparseArray(arr_data).copy()

        def setitem():
            arr[5] = 3

        def setslice():
            arr[1:5] = 2

        with pytest.raises(TypeError, match="assignment via setitem"):
            setitem()

        with pytest.raises(TypeError, match="assignment via setitem"):
            setslice()


class TestTake:
    def test_take_scalar_raises(self, arr):
        msg = "'indices' must be an array, not a scalar '2'."
        with pytest.raises(ValueError, match=msg):
            arr.take(2)

    def test_take(self, arr_data, arr):
        exp = SparseArray(np.take(arr_data, [2, 3]))
        tm.assert_sp_array_equal(arr.take([2, 3]), exp)

        exp = SparseArray(np.take(arr_data, [0, 1, 2]))
        tm.assert_sp_array_equal(arr.take([0, 1, 2]), exp)

    def test_take_all_empty(self):
        sparse = pd.array([0, 0], dtype=SparseDtype("int64"))
        result = sparse.take([0, 1], allow_fill=True, fill_value=np.nan)
        tm.assert_sp_array_equal(sparse, result)

    def test_take_different_fill_value(self):
        # Take with a different fill value shouldn't overwrite the original
        sparse = pd.array([0.0], dtype=SparseDtype("float64", fill_value=0.0))
        result = sparse.take([0, -1], allow_fill=True, fill_value=np.nan)
        expected = pd.array([0, np.nan], dtype=sparse.dtype)
        tm.assert_sp_array_equal(expected, result)

    def test_take_fill_value(self):
        data = np.array([1, np.nan, 0, 3, 0])
        sparse = SparseArray(data, fill_value=0)

        exp = SparseArray(np.take(data, [0]), fill_value=0)
        tm.assert_sp_array_equal(sparse.take([0]), exp)

        exp = SparseArray(np.take(data, [1, 3, 4]), fill_value=0)
        tm.assert_sp_array_equal(sparse.take([1, 3, 4]), exp)

    def test_take_negative(self, arr_data, arr):
        exp = SparseArray(np.take(arr_data, [-1]))
        tm.assert_sp_array_equal(arr.take([-1]), exp)

        exp = SparseArray(np.take(arr_data, [-4, -3, -2]))
        tm.assert_sp_array_equal(arr.take([-4, -3, -2]), exp)

    def test_bad_take(self, arr):
        with pytest.raises(IndexError, match="bounds"):
            arr.take([11])

    def test_take_filling(self):
        # similar tests as GH 12631
        sparse = SparseArray([np.nan, np.nan, 1, np.nan, 4])
        result = sparse.take(np.array([1, 0, -1]))
        expected = SparseArray([np.nan, np.nan, 4])
        tm.assert_sp_array_equal(result, expected)

        # TODO: actionable?
        # XXX: test change: fill_value=True -> allow_fill=True
        result = sparse.take(np.array([1, 0, -1]), allow_fill=True)
        expected = SparseArray([np.nan, np.nan, np.nan])
        tm.assert_sp_array_equal(result, expected)

        # allow_fill=False
        result = sparse.take(np.array([1, 0, -1]), allow_fill=False, fill_value=True)
        expected = SparseArray([np.nan, np.nan, 4])
        tm.assert_sp_array_equal(result, expected)

        msg = "Invalid value in 'indices'"
        with pytest.raises(ValueError, match=msg):
            sparse.take(np.array([1, 0, -2]), allow_fill=True)

        with pytest.raises(ValueError, match=msg):
            sparse.take(np.array([1, 0, -5]), allow_fill=True)

        msg = "out of bounds value in 'indices'"
        with pytest.raises(IndexError, match=msg):
            sparse.take(np.array([1, -6]))
        with pytest.raises(IndexError, match=msg):
            sparse.take(np.array([1, 5]))
        with pytest.raises(IndexError, match=msg):
            sparse.take(np.array([1, 5]), allow_fill=True)

    def test_take_filling_fill_value(self):
        # same tests as GH#12631
        sparse = SparseArray([np.nan, 0, 1, 0, 4], fill_value=0)
        result = sparse.take(np.array([1, 0, -1]))
        expected = SparseArray([0, np.nan, 4], fill_value=0)
        tm.assert_sp_array_equal(result, expected)

        # fill_value
        result = sparse.take(np.array([1, 0, -1]), allow_fill=True)
        # TODO: actionable?
        # XXX: behavior change.
        # the old way of filling self.fill_value doesn't follow EA rules.
        # It's supposed to be self.dtype.na_value (nan in this case)
        expected = SparseArray([0, np.nan, np.nan], fill_value=0)
        tm.assert_sp_array_equal(result, expected)

        # allow_fill=False
        result = sparse.take(np.array([1, 0, -1]), allow_fill=False, fill_value=True)
        expected = SparseArray([0, np.nan, 4], fill_value=0)
        tm.assert_sp_array_equal(result, expected)

        msg = "Invalid value in 'indices'."
        with pytest.raises(ValueError, match=msg):
            sparse.take(np.array([1, 0, -2]), allow_fill=True)
        with pytest.raises(ValueError, match=msg):
            sparse.take(np.array([1, 0, -5]), allow_fill=True)

        msg = "out of bounds value in 'indices'"
        with pytest.raises(IndexError, match=msg):
            sparse.take(np.array([1, -6]))
        with pytest.raises(IndexError, match=msg):
            sparse.take(np.array([1, 5]))
        with pytest.raises(IndexError, match=msg):
            sparse.take(np.array([1, 5]), fill_value=True)

    @pytest.mark.parametrize("kind", ["block", "integer"])
    def test_take_filling_all_nan(self, kind):
        sparse = SparseArray([np.nan, np.nan, np.nan, np.nan, np.nan], kind=kind)
        result = sparse.take(np.array([1, 0, -1]))
        expected = SparseArray([np.nan, np.nan, np.nan], kind=kind)
        tm.assert_sp_array_equal(result, expected)

        result = sparse.take(np.array([1, 0, -1]), fill_value=True)
        expected = SparseArray([np.nan, np.nan, np.nan], kind=kind)
        tm.assert_sp_array_equal(result, expected)

        msg = "out of bounds value in 'indices'"
        with pytest.raises(IndexError, match=msg):
            sparse.take(np.array([1, -6]))
        with pytest.raises(IndexError, match=msg):
            sparse.take(np.array([1, 5]))
        with pytest.raises(IndexError, match=msg):
            sparse.take(np.array([1, 5]), fill_value=True)


class TestWhere:
    def test_where_retain_fill_value(self):
        # GH#45691 don't lose fill_value on _where
        arr = SparseArray([np.nan, 1.0], fill_value=0)

        mask = np.array([True, False])

        res = arr._where(~mask, 1)
        exp = SparseArray([1, 1.0], fill_value=0)
        tm.assert_sp_array_equal(res, exp)

        ser = pd.Series(arr)
        res = ser.where(~mask, 1)
        tm.assert_series_equal(res, pd.Series(exp))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\arrays\integer\test_dtypes.py ===
import numpy as np
import pytest

from pandas.core.dtypes.generic import ABCIndex

import pandas as pd
import pandas._testing as tm
from pandas.core.arrays.integer import (
    Int8Dtype,
    UInt32Dtype,
)


def test_dtypes(dtype):
    # smoke tests on auto dtype construction

    if dtype.is_signed_integer:
        assert np.dtype(dtype.type).kind == "i"
    else:
        assert np.dtype(dtype.type).kind == "u"
    assert dtype.name is not None


@pytest.mark.parametrize("op", ["sum", "min", "max", "prod"])
def test_preserve_dtypes(op):
    # for ops that enable (mean would actually work here
    # but generally it is a float return value)
    df = pd.DataFrame(
        {
            "A": ["a", "b", "b"],
            "B": [1, None, 3],
            "C": pd.array([1, None, 3], dtype="Int64"),
        }
    )

    # op
    result = getattr(df.C, op)()
    if op in {"sum", "prod", "min", "max"}:
        assert isinstance(result, np.int64)
    else:
        assert isinstance(result, int)

    # groupby
    result = getattr(df.groupby("A"), op)()

    expected = pd.DataFrame(
        {"B": np.array([1.0, 3.0]), "C": pd.array([1, 3], dtype="Int64")},
        index=pd.Index(["a", "b"], name="A"),
    )
    tm.assert_frame_equal(result, expected)


def test_astype_nansafe():
    # see gh-22343
    arr = pd.array([np.nan, 1, 2], dtype="Int8")
    msg = "cannot convert NA to integer"

    with pytest.raises(ValueError, match=msg):
        arr.astype("uint32")


@pytest.mark.parametrize("dropna", [True, False])
def test_construct_index(all_data, dropna):
    # ensure that we do not coerce to different Index dtype or non-index

    all_data = all_data[:10]
    if dropna:
        other = np.array(all_data[~all_data.isna()])
    else:
        other = all_data

    result = pd.Index(pd.array(other, dtype=all_data.dtype))
    expected = pd.Index(other, dtype=all_data.dtype)
    assert all_data.dtype == expected.dtype  # dont coerce to object

    tm.assert_index_equal(result, expected)


@pytest.mark.parametrize("dropna", [True, False])
def test_astype_index(all_data, dropna):
    # as an int/uint index to Index

    all_data = all_data[:10]
    if dropna:
        other = all_data[~all_data.isna()]
    else:
        other = all_data

    dtype = all_data.dtype
    idx = pd.Index(np.array(other))
    assert isinstance(idx, ABCIndex)

    result = idx.astype(dtype)
    expected = idx.astype(object).astype(dtype)
    tm.assert_index_equal(result, expected)


def test_astype(all_data):
    all_data = all_data[:10]

    ints = all_data[~all_data.isna()]
    mixed = all_data
    dtype = Int8Dtype()

    # coerce to same type - ints
    s = pd.Series(ints)
    result = s.astype(all_data.dtype)
    expected = pd.Series(ints)
    tm.assert_series_equal(result, expected)

    # coerce to same other - ints
    s = pd.Series(ints)
    result = s.astype(dtype)
    expected = pd.Series(ints, dtype=dtype)
    tm.assert_series_equal(result, expected)

    # coerce to same numpy_dtype - ints
    s = pd.Series(ints)
    result = s.astype(all_data.dtype.numpy_dtype)
    expected = pd.Series(ints._data.astype(all_data.dtype.numpy_dtype))
    tm.assert_series_equal(result, expected)

    # coerce to same type - mixed
    s = pd.Series(mixed)
    result = s.astype(all_data.dtype)
    expected = pd.Series(mixed)
    tm.assert_series_equal(result, expected)

    # coerce to same other - mixed
    s = pd.Series(mixed)
    result = s.astype(dtype)
    expected = pd.Series(mixed, dtype=dtype)
    tm.assert_series_equal(result, expected)

    # coerce to same numpy_dtype - mixed
    s = pd.Series(mixed)
    msg = "cannot convert NA to integer"
    with pytest.raises(ValueError, match=msg):
        s.astype(all_data.dtype.numpy_dtype)

    # coerce to object
    s = pd.Series(mixed)
    result = s.astype("object")
    expected = pd.Series(np.asarray(mixed, dtype=object))
    tm.assert_series_equal(result, expected)


def test_astype_copy():
    arr = pd.array([1, 2, 3, None], dtype="Int64")
    orig = pd.array([1, 2, 3, None], dtype="Int64")

    # copy=True -> ensure both data and mask are actual copies
    result = arr.astype("Int64", copy=True)
    assert result is not arr
    assert not tm.shares_memory(result, arr)
    result[0] = 10
    tm.assert_extension_array_equal(arr, orig)
    result[0] = pd.NA
    tm.assert_extension_array_equal(arr, orig)

    # copy=False
    result = arr.astype("Int64", copy=False)
    assert result is arr
    assert np.shares_memory(result._data, arr._data)
    assert np.shares_memory(result._mask, arr._mask)
    result[0] = 10
    assert arr[0] == 10
    result[0] = pd.NA
    assert arr[0] is pd.NA

    # astype to different dtype -> always needs a copy -> even with copy=False
    # we need to ensure that also the mask is actually copied
    arr = pd.array([1, 2, 3, None], dtype="Int64")
    orig = pd.array([1, 2, 3, None], dtype="Int64")

    result = arr.astype("Int32", copy=False)
    assert not tm.shares_memory(result, arr)
    result[0] = 10
    tm.assert_extension_array_equal(arr, orig)
    result[0] = pd.NA
    tm.assert_extension_array_equal(arr, orig)


def test_astype_to_larger_numpy():
    a = pd.array([1, 2], dtype="Int32")
    result = a.astype("int64")
    expected = np.array([1, 2], dtype="int64")
    tm.assert_numpy_array_equal(result, expected)

    a = pd.array([1, 2], dtype="UInt32")
    result = a.astype("uint64")
    expected = np.array([1, 2], dtype="uint64")
    tm.assert_numpy_array_equal(result, expected)


@pytest.mark.parametrize("dtype", [Int8Dtype(), "Int8", UInt32Dtype(), "UInt32"])
def test_astype_specific_casting(dtype):
    s = pd.Series([1, 2, 3], dtype="Int64")
    result = s.astype(dtype)
    expected = pd.Series([1, 2, 3], dtype=dtype)
    tm.assert_series_equal(result, expected)

    s = pd.Series([1, 2, 3, None], dtype="Int64")
    result = s.astype(dtype)
    expected = pd.Series([1, 2, 3, None], dtype=dtype)
    tm.assert_series_equal(result, expected)


def test_astype_floating():
    arr = pd.array([1, 2, None], dtype="Int64")
    result = arr.astype("Float64")
    expected = pd.array([1.0, 2.0, None], dtype="Float64")
    tm.assert_extension_array_equal(result, expected)


def test_astype_dt64():
    # GH#32435
    arr = pd.array([1, 2, 3, pd.NA]) * 10**9

    result = arr.astype("datetime64[ns]")

    expected = np.array([1, 2, 3, "NaT"], dtype="M8[s]").astype("M8[ns]")
    tm.assert_numpy_array_equal(result, expected)


def test_construct_cast_invalid(dtype):
    msg = "cannot safely"
    arr = [1.2, 2.3, 3.7]
    with pytest.raises(TypeError, match=msg):
        pd.array(arr, dtype=dtype)

    with pytest.raises(TypeError, match=msg):
        pd.Series(arr).astype(dtype)

    arr = [1.2, 2.3, 3.7, np.nan]
    with pytest.raises(TypeError, match=msg):
        pd.array(arr, dtype=dtype)

    with pytest.raises(TypeError, match=msg):
        pd.Series(arr).astype(dtype)


@pytest.mark.parametrize("in_series", [True, False])
def test_to_numpy_na_nan(in_series):
    a = pd.array([0, 1, None], dtype="Int64")
    if in_series:
        a = pd.Series(a)

    result = a.to_numpy(dtype="float64", na_value=np.nan)
    expected = np.array([0.0, 1.0, np.nan], dtype="float64")
    tm.assert_numpy_array_equal(result, expected)

    result = a.to_numpy(dtype="int64", na_value=-1)
    expected = np.array([0, 1, -1], dtype="int64")
    tm.assert_numpy_array_equal(result, expected)

    result = a.to_numpy(dtype="bool", na_value=False)
    expected = np.array([False, True, False], dtype="bool")
    tm.assert_numpy_array_equal(result, expected)


@pytest.mark.parametrize("in_series", [True, False])
@pytest.mark.parametrize("dtype", ["int32", "int64", "bool"])
def test_to_numpy_dtype(dtype, in_series):
    a = pd.array([0, 1], dtype="Int64")
    if in_series:
        a = pd.Series(a)

    result = a.to_numpy(dtype=dtype)
    expected = np.array([0, 1], dtype=dtype)
    tm.assert_numpy_array_equal(result, expected)


@pytest.mark.parametrize("dtype", ["int64", "bool"])
def test_to_numpy_na_raises(dtype):
    a = pd.array([0, 1, None], dtype="Int64")
    with pytest.raises(ValueError, match=dtype):
        a.to_numpy(dtype=dtype)


def test_astype_str(using_infer_string):
    a = pd.array([1, 2, None], dtype="Int64")

    if using_infer_string:
        expected = pd.array(["1", "2", None], dtype=pd.StringDtype(na_value=np.nan))

        tm.assert_extension_array_equal(a.astype(str), expected)
        tm.assert_extension_array_equal(a.astype("str"), expected)
    else:
        expected = np.array(["1", "2", "<NA>"], dtype=f"{tm.ENDIAN}U21")

        tm.assert_numpy_array_equal(a.astype(str), expected)
        tm.assert_numpy_array_equal(a.astype("str"), expected)


def test_astype_boolean():
    # https://github.com/pandas-dev/pandas/issues/31102
    a = pd.array([1, 0, -1, 2, None], dtype="Int64")
    result = a.astype("boolean")
    expected = pd.array([True, False, True, True, None], dtype="boolean")
    tm.assert_extension_array_equal(result, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\groupby\methods\test_describe.py ===
import numpy as np
import pytest

import pandas as pd
from pandas import (
    DataFrame,
    Index,
    MultiIndex,
    Series,
    Timestamp,
    date_range,
)
import pandas._testing as tm


def test_apply_describe_bug(multiindex_dataframe_random_data):
    grouped = multiindex_dataframe_random_data.groupby(level="first")
    grouped.describe()  # it works!


def test_series_describe_multikey():
    ts = Series(
        np.arange(10, dtype=np.float64), index=date_range("2020-01-01", periods=10)
    )
    grouped = ts.groupby([lambda x: x.year, lambda x: x.month])
    result = grouped.describe()
    tm.assert_series_equal(result["mean"], grouped.mean(), check_names=False)
    tm.assert_series_equal(result["std"], grouped.std(), check_names=False)
    tm.assert_series_equal(result["min"], grouped.min(), check_names=False)


def test_series_describe_single():
    ts = Series(
        np.arange(10, dtype=np.float64), index=date_range("2020-01-01", periods=10)
    )
    grouped = ts.groupby(lambda x: x.month)
    result = grouped.apply(lambda x: x.describe())
    expected = grouped.describe().stack(future_stack=True)
    tm.assert_series_equal(result, expected)


@pytest.mark.parametrize("keys", ["key1", ["key1", "key2"]])
def test_series_describe_as_index(as_index, keys):
    # GH#49256
    df = DataFrame(
        {
            "key1": ["one", "two", "two", "three", "two"],
            "key2": ["one", "two", "two", "three", "two"],
            "foo2": [1, 2, 4, 4, 6],
        }
    )
    gb = df.groupby(keys, as_index=as_index)["foo2"]
    result = gb.describe()
    expected = DataFrame(
        {
            "key1": ["one", "three", "two"],
            "count": [1.0, 1.0, 3.0],
            "mean": [1.0, 4.0, 4.0],
            "std": [np.nan, np.nan, 2.0],
            "min": [1.0, 4.0, 2.0],
            "25%": [1.0, 4.0, 3.0],
            "50%": [1.0, 4.0, 4.0],
            "75%": [1.0, 4.0, 5.0],
            "max": [1.0, 4.0, 6.0],
        }
    )
    if len(keys) == 2:
        expected.insert(1, "key2", expected["key1"])
    if as_index:
        expected = expected.set_index(keys)
    tm.assert_frame_equal(result, expected)


def test_frame_describe_multikey(tsframe, using_infer_string):
    grouped = tsframe.groupby([lambda x: x.year, lambda x: x.month])
    result = grouped.describe()
    desc_groups = []
    for col in tsframe:
        group = grouped[col].describe()
        # GH 17464 - Remove duplicate MultiIndex levels
        group_col = MultiIndex(
            levels=[Index([col], dtype=tsframe.columns.dtype), group.columns],
            codes=[[0] * len(group.columns), range(len(group.columns))],
        )
        group = DataFrame(group.values, columns=group_col, index=group.index)
        desc_groups.append(group)
    expected = pd.concat(desc_groups, axis=1)
    tm.assert_frame_equal(result, expected)

    # remainder of the tests fails with string dtype but is testing deprecated behaviour
    if using_infer_string:
        return

    msg = "DataFrame.groupby with axis=1 is deprecated"
    with tm.assert_produces_warning(FutureWarning, match=msg):
        groupedT = tsframe.groupby({"A": 0, "B": 0, "C": 1, "D": 1}, axis=1)
    result = groupedT.describe()
    expected = tsframe.describe().T
    # reverting the change from https://github.com/pandas-dev/pandas/pull/35441/
    expected.index = MultiIndex(
        levels=[[0, 1], expected.index],
        codes=[[0, 0, 1, 1], range(len(expected.index))],
    )
    tm.assert_frame_equal(result, expected)


def test_frame_describe_tupleindex():
    # GH 14848 - regression from 0.19.0 to 0.19.1
    df1 = DataFrame(
        {
            "x": [1, 2, 3, 4, 5] * 3,
            "y": [10, 20, 30, 40, 50] * 3,
            "z": [100, 200, 300, 400, 500] * 3,
        }
    )
    df1["k"] = [(0, 0, 1), (0, 1, 0), (1, 0, 0)] * 5
    df2 = df1.rename(columns={"k": "key"})
    msg = "Names should be list-like for a MultiIndex"
    with pytest.raises(ValueError, match=msg):
        df1.groupby("k").describe()
    with pytest.raises(ValueError, match=msg):
        df2.groupby("key").describe()


def test_frame_describe_unstacked_format():
    # GH 4792
    prices = {
        Timestamp("2011-01-06 10:59:05", tz=None): 24990,
        Timestamp("2011-01-06 12:43:33", tz=None): 25499,
        Timestamp("2011-01-06 12:54:09", tz=None): 25499,
    }
    volumes = {
        Timestamp("2011-01-06 10:59:05", tz=None): 1500000000,
        Timestamp("2011-01-06 12:43:33", tz=None): 5000000000,
        Timestamp("2011-01-06 12:54:09", tz=None): 100000000,
    }
    df = DataFrame({"PRICE": prices, "VOLUME": volumes})
    result = df.groupby("PRICE").VOLUME.describe()
    data = [
        df[df.PRICE == 24990].VOLUME.describe().values.tolist(),
        df[df.PRICE == 25499].VOLUME.describe().values.tolist(),
    ]
    expected = DataFrame(
        data,
        index=Index([24990, 25499], name="PRICE"),
        columns=["count", "mean", "std", "min", "25%", "50%", "75%", "max"],
    )
    tm.assert_frame_equal(result, expected)


@pytest.mark.filterwarnings(
    "ignore:"
    "indexing past lexsort depth may impact performance:"
    "pandas.errors.PerformanceWarning"
)
@pytest.mark.parametrize("as_index", [True, False])
@pytest.mark.parametrize("keys", [["a1"], ["a1", "a2"]])
def test_describe_with_duplicate_output_column_names(as_index, keys):
    # GH 35314
    df = DataFrame(
        {
            "a1": [99, 99, 99, 88, 88, 88],
            "a2": [99, 99, 99, 88, 88, 88],
            "b": [1, 2, 3, 4, 5, 6],
            "c": [10, 20, 30, 40, 50, 60],
        },
        columns=["a1", "a2", "b", "b"],
        copy=False,
    )
    if keys == ["a1"]:
        df = df.drop(columns="a2")

    expected = (
        DataFrame.from_records(
            [
                ("b", "count", 3.0, 3.0),
                ("b", "mean", 5.0, 2.0),
                ("b", "std", 1.0, 1.0),
                ("b", "min", 4.0, 1.0),
                ("b", "25%", 4.5, 1.5),
                ("b", "50%", 5.0, 2.0),
                ("b", "75%", 5.5, 2.5),
                ("b", "max", 6.0, 3.0),
                ("b", "count", 3.0, 3.0),
                ("b", "mean", 5.0, 2.0),
                ("b", "std", 1.0, 1.0),
                ("b", "min", 4.0, 1.0),
                ("b", "25%", 4.5, 1.5),
                ("b", "50%", 5.0, 2.0),
                ("b", "75%", 5.5, 2.5),
                ("b", "max", 6.0, 3.0),
            ],
        )
        .set_index([0, 1])
        .T
    )
    expected.columns.names = [None, None]
    if len(keys) == 2:
        expected.index = MultiIndex(
            levels=[[88, 99], [88, 99]], codes=[[0, 1], [0, 1]], names=["a1", "a2"]
        )
    else:
        expected.index = Index([88, 99], name="a1")

    if not as_index:
        expected = expected.reset_index()

    result = df.groupby(keys, as_index=as_index).describe()

    tm.assert_frame_equal(result, expected)


def test_describe_duplicate_columns():
    # GH#50806
    df = DataFrame([[0, 1, 2, 3]])
    df.columns = [0, 1, 2, 0]
    gb = df.groupby(df[1])
    result = gb.describe(percentiles=[])

    columns = ["count", "mean", "std", "min", "50%", "max"]
    frames = [
        DataFrame([[1.0, val, np.nan, val, val, val]], index=[1], columns=columns)
        for val in (0.0, 2.0, 3.0)
    ]
    expected = pd.concat(frames, axis=1)
    expected.columns = MultiIndex(
        levels=[[0, 2], columns],
        codes=[6 * [0] + 6 * [1] + 6 * [0], 3 * list(range(6))],
    )
    expected.index.names = [1]
    tm.assert_frame_equal(result, expected)


class TestGroupByNonCythonPaths:
    # GH#5610 non-cython calls should not include the grouper
    # Tests for code not expected to go through cython paths.

    @pytest.fixture
    def df(self):
        df = DataFrame(
            [[1, 2, "foo"], [1, np.nan, "bar"], [3, np.nan, "baz"]],
            columns=["A", "B", "C"],
        )
        return df

    @pytest.fixture
    def gb(self, df):
        gb = df.groupby("A")
        return gb

    @pytest.fixture
    def gni(self, df):
        gni = df.groupby("A", as_index=False)
        return gni

    def test_describe(self, df, gb, gni):
        # describe
        expected_index = Index([1, 3], name="A")
        expected_col = MultiIndex(
            levels=[["B"], ["count", "mean", "std", "min", "25%", "50%", "75%", "max"]],
            codes=[[0] * 8, list(range(8))],
        )
        expected = DataFrame(
            [
                [1.0, 2.0, np.nan, 2.0, 2.0, 2.0, 2.0, 2.0],
                [0.0, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan],
            ],
            index=expected_index,
            columns=expected_col,
        )
        result = gb.describe()
        tm.assert_frame_equal(result, expected)

        expected = expected.reset_index()
        result = gni.describe()
        tm.assert_frame_equal(result, expected)


@pytest.mark.parametrize("dtype", [int, float, object])
@pytest.mark.parametrize(
    "kwargs",
    [
        {"percentiles": [0.10, 0.20, 0.30], "include": "all", "exclude": None},
        {"percentiles": [0.10, 0.20, 0.30], "include": None, "exclude": ["int"]},
        {"percentiles": [0.10, 0.20, 0.30], "include": ["int"], "exclude": None},
    ],
)
def test_groupby_empty_dataset(dtype, kwargs):
    # GH#41575
    df = DataFrame([[1, 2, 3]], columns=["A", "B", "C"], dtype=dtype)
    df["B"] = df["B"].astype(int)
    df["C"] = df["C"].astype(float)

    result = df.iloc[:0].groupby("A").describe(**kwargs)
    expected = df.groupby("A").describe(**kwargs).reset_index(drop=True).iloc[:0]
    tm.assert_frame_equal(result, expected)

    result = df.iloc[:0].groupby("A").B.describe(**kwargs)
    expected = df.groupby("A").B.describe(**kwargs).reset_index(drop=True).iloc[:0]
    expected.index = Index([], dtype=df.columns.dtype)
    tm.assert_frame_equal(result, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\matrices\tests\test_fflu.py ===
from sympy.polys.matrices import DomainMatrix, DM
from sympy.polys.domains import ZZ, QQ
from sympy import Matrix
import pytest


FFLU_EXAMPLES = [
    (
        'zz_2x3',
        DM([[1, 2, 3], [4, 5, 6]], ZZ),
        DM([[1, 0], [0, 1]], ZZ),
        DM([[1, 0], [4, -3]], ZZ),
        DM([[1, 0], [0, -3]], ZZ),
        DM([[1, 2, 3], [0, -3, -6]], ZZ),
    ),

    (
        'zz_2x2',
        DM([[4, 3], [6, 3]], ZZ),
        DM([[1, 0], [0, 1]], ZZ),
        DM([[1, 0], [6, -6]], ZZ),
        DM([[4, 0], [0, -3]], ZZ),
        DM([[4, 3], [0, -3]], ZZ),
    ),

    (
        'zz_3x2',
        DM([[1, 2], [3, 4], [5, 6]], ZZ),
        DM([[1, 0, 0], [0, 1, 0], [0, 0, 1]], ZZ),
        DM([[1, 0, 0], [3, 1, 0], [5, 2, 1]], ZZ),
        DM([[1, 0], [0, -2]], ZZ),
        DM([[1, 2], [0, -2], [0, 0]], ZZ),
    ),

    (
        'zz_3x3',
        DM([[1, 2, 3], [4, 5, 6], [7, 8, 9]], ZZ),
        DM([[1, 0, 0], [0, 1, 0], [0, 0, 1]], ZZ),
        DM([[1, 0, 0], [4, 1, 0], [7, 2, 1]], ZZ),
        DM([[1, 0, 0], [0, -3, 0], [0, 0, 0]], ZZ),
        DM([[1, 2, 3], [0, -3, -6], [0, 0, 0]], ZZ),
    ),

    (
        'zz_zero',
        DM([[0, 0, 0], [0, 0, 0], [0, 0, 0]], ZZ),
        DM([[1, 0, 0], [0, 1, 0], [0, 0, 1]], ZZ),
        DM([[1, 0, 0], [0, 1, 0], [0, 0, 1]], ZZ),
        DM([[0, 0, 0], [0, 0, 0], [0, 0, 0]], ZZ),
        DM([[0, 0, 0], [0, 0, 0], [0, 0, 0]], ZZ),
    ),

    (
        'zz_empty',
        DM([], ZZ),
        DM([], ZZ),
        DM([], ZZ),
        DM([], ZZ),
        DM([], ZZ),
    ),

    (
        'zz_empty_0x2',
        DomainMatrix([], (0, 2), ZZ),
        DomainMatrix([], (0, 0), ZZ),
        DomainMatrix([], (0, 0), ZZ),
        DomainMatrix([], (0, 0), ZZ),
        DomainMatrix([], (0, 2), ZZ)
    ),

    (

        'zz_empty_2x0',
        DomainMatrix([[], []], (2, 0), ZZ),
        DomainMatrix.eye((2, 2), ZZ),
        DomainMatrix.eye((2, 2), ZZ),
        DomainMatrix.eye((2, 2), ZZ),
        DomainMatrix([[], []], (2, 0), ZZ)

    ),

    (
        'zz_negative',
        DM([[-1, -2], [-3, -4]], ZZ),
        DM([[1, 0], [0, 1]], ZZ),
        DM([[-1, 0], [-3, -2]], ZZ),
        DM([[-1, 0], [0, 2]], ZZ),
        DM([[-1, -2], [0, -2]], ZZ),
    ),

    (
        'zz_mixed_signs',
        DM([[1, -2], [-3, 4]], ZZ),
        DM([[1, 0], [0, 1]], ZZ),
        DM([[1, 0], [-3, 1]], ZZ),
        DM([[1, 0], [0, -2]], ZZ),
        DM([[1, -2], [0, -2]], ZZ),
    ),

    (
        'zz_upper_triangular',
        DM([[1, 2, 3], [0, 4, 5], [0, 0, 6]], ZZ),
        DM([[1, 0, 0], [0, 1, 0], [0, 0, 1]], ZZ),
        DM([[1, 0, 0], [0, 4, 0], [0, 0, 24]], ZZ),
        DM([[1, 0, 0], [0, 4, 0], [0, 0, 96]], ZZ),
        DM([[1, 2, 3], [0, 4, 5], [0, 0, 24]], ZZ),
    ),

    (
        'zz_lower_triangular',
        DM([[1, 0, 0], [2, 3, 0], [4, 5, 6]], ZZ),
        DM([[1, 0, 0], [0, 1, 0], [0, 0, 1]], ZZ),
        DM([[1, 0, 0], [2, 3, 0], [4, 5, 18]], ZZ),
        DM([[1, 0, 0], [0, 3, 0], [0, 0, 54]], ZZ),
        DM([[1, 0, 0], [0, 3, 0], [0, 0, 18]], ZZ),
    ),

    (
        'zz_diagonal',
        DM([[2, 0, 0], [0, 3, 0], [0, 0, 4]], ZZ),
        DM([[1, 0, 0], [0, 1, 0], [0, 0, 1]], ZZ),
        DM([[2, 0, 0], [0, 6, 0], [0, 0, 24]], ZZ),
        DM([[2, 0, 0], [0, 12, 0], [0, 0, 144]], ZZ),
        DM([[2, 0, 0], [0, 6, 0], [0, 0, 24]], ZZ)

    ),

    (
        'rank_deficient_3x3',
        DM([[1, 2, 3], [2, 4, 6], [3, 6, 9]], ZZ),
        DM([[1, 0, 0], [0, 1, 0], [0, 0, 1]], ZZ),
        DM([[1, 0, 0], [2, 1, 0], [3, 0, 1]], ZZ),
        DM([[1, 0, 0], [0, 0, 0], [0, 0, 0]], ZZ),
        DM([[1, 2, 3], [0, 0, 0], [0, 0, 0]], ZZ),
    ),

    (
        'zz_1x1',
        DM([[5]], ZZ),
        DM([[1]], ZZ),
        DM([[5]], ZZ),
        DM([[5]], ZZ),
        DM([[5]], ZZ),
    ),

    (
        'zz_nx1_2rows',
        DM([[81], [54]], ZZ),
        DM([[1, 0], [0, 1]], ZZ),
        DM([[81, 0], [54, 81]], ZZ),
        DM([[81, 0], [0, 81]], ZZ),
        DM([[81], [0]], ZZ),
    ),

    (
        'zz_nx2_3rows',
        DM([[2, 7], [7, 45], [25, 84]], ZZ),
        DM([[1, 0, 0], [0, 1, 0], [0, 0, 1]], ZZ),
        DM([[2, 0, 0], [7, 82, 0], [25, 41, 41]], ZZ),
        DM([[2, 0, 0], [0, 82, 0], [0, 0, 41]], ZZ),
        DM([[2, 7], [0, 82], [0, 0]], ZZ),
    ),

    (

        'zz_1x2',
        DM([[0, 28]], ZZ),
        DM([[1]], ZZ),
        DM([[28]], ZZ),
        DM([[28]], ZZ),
        DM([[0, 28]], ZZ)
    ),

    (
        'zz_nx3_4rows',
        DM([[84, 30, 9], [20, 59, 13], [53, 46, 81], [63, 48, 29]], ZZ),
        DM([[1, 0, 0, 0], [0, 1, 0, 0], [0, 0, 1, 0], [0, 0, 0, 1]], ZZ),
        DM([[84, 0, 0, 0], [20, 365904, 0, 0], [53, 303411, 303411, 0], [63, 303411, 303411, 303411]], ZZ),
        DM([[84, 0, 0, 0], [0, 365904, 0, 0], [0, 0, 1321658316, 0], [0, 0, 0, 303411]], ZZ),
        DM([[84, 30, 9], [0, 365904, 13], [0, 0, 1321658316], [0, 0, 0]], ZZ),
    ),

    (
        'fflu_row_swap',
        DM([[0, 1, 2], [3, 4, 5], [6, 7, 8]], ZZ),
        DM([[0, 1, 0], [1, 0, 0], [0, 0, 1]], ZZ),
        DM([[3, 0, 0], [0, 3, 0], [6, -3, 1]], ZZ),
        DM([[3, 0, 0], [0, 9, 0], [0, 0, 3]], ZZ),
        DM([[3, 4, 5], [0, 3, 6], [0, 0, 0]], ZZ)
    ),
]


def _check_fflu(A, P, L, D, U):
    P_field = P.to_field().to_dense()
    L_field = L.to_field().to_dense()
    D_field = D.to_field().to_dense()
    U_field = U.to_field().to_dense()
    m, n = A.shape
    assert P_field.shape == (m, m)
    assert L_field.shape == (m, m)
    assert D_field.shape == (m, m)
    assert U_field.shape == (m, n)
    assert L_field.is_lower
    assert D_field.is_diagonal
    di, d = D.inv_den()
    assert P.matmul(A).rmul(d) == L.matmul(di).matmul(U)
    assert U_field.is_upper


def _to_DM(A, ans):
    if isinstance(A, DomainMatrix):
        return A
    elif isinstance(A, Matrix):
        return A.to_DM(ans.domain)
    return DomainMatrix(A.to_list(), A.shape, A.domain)


def _check_fflu_result(result, A, P_ans, L_ans, D_ans, U_ans):
    P, L, D, U = result
    P = _to_DM(P, P_ans)
    L = _to_DM(L, L_ans)
    D = _to_DM(D, D_ans)
    U = _to_DM(U, U_ans)
    A = _to_DM(A, P_ans)
    m, n = A.shape
    assert P.shape == (m, m)
    assert L.shape == (m, m)
    assert D.shape == (m, m)
    assert U.shape == (m, n)
    assert L.is_lower
    assert D.is_diagonal
    di, d = D.inv_den()
    assert P.matmul(A).rmul(d) == L.matmul(di).matmul(U)
    assert U.is_upper


@pytest.mark.parametrize('name, A, P_ans, L_ans, D_ans, U_ans', FFLU_EXAMPLES)
def test_dm_dense_fflu(name, A, P_ans, L_ans, D_ans, U_ans):
    A = A.to_dense()
    _check_fflu_result(A.fflu(), A, P_ans, L_ans, D_ans, U_ans)


@pytest.mark.parametrize('name, A, P_ans, L_ans, D_ans, U_ans', FFLU_EXAMPLES)
def test_dm_sparse_fflu(name, A, P_ans, L_ans, D_ans, U_ans):
    A = A.to_sparse()
    _check_fflu_result(A.fflu(), A, P_ans, L_ans, D_ans, U_ans)


@pytest.mark.parametrize('name, A, P_ans, L_ans, D_ans, U_ans', FFLU_EXAMPLES)
def test_ddm_fflu(name, A, P_ans, L_ans, D_ans, U_ans):
    A = A.to_ddm()
    _check_fflu_result(A.fflu(), A, P_ans, L_ans, D_ans, U_ans)


@pytest.mark.parametrize('name, A, P_ans, L_ans, D_ans, U_ans', FFLU_EXAMPLES)
def test_sdm_fflu(name, A, P_ans, L_ans, D_ans, U_ans):
    A = A.to_sdm()
    _check_fflu_result(A.fflu(), A, P_ans, L_ans, D_ans, U_ans)


@pytest.mark.parametrize('name, A, P_ans, L_ans, D_ans, U_ans', FFLU_EXAMPLES)
def test_dfm_fflu(name, A, P_ans, L_ans, D_ans, U_ans):
    pytest.importorskip('flint')
    if A.domain not in (ZZ, QQ) and not A.domain.is_FF:
        pytest.skip("Domain not supported by DFM")
    A = A.to_dfm()
    _check_fflu_result(A.fflu(), A, P_ans, L_ans, D_ans, U_ans)


def test_fflu_empty_matrix():
    A = DomainMatrix([], (0, 0), ZZ)
    P, L, D, U = A.fflu()
    assert P.shape == (0, 0)
    assert L.shape == (0, 0)
    assert D.shape == (0, 0)
    assert U.shape == (0, 0)


def test_fflu_properties():
    A = DomainMatrix([[ZZ(1), ZZ(2)], [ZZ(3), ZZ(4)]], (2, 2), ZZ)
    P, L, D, U = A.fflu()
    assert P.shape == (2, 2)
    assert L.shape == (2, 2)
    assert D.shape == (2, 2)
    assert U.shape == (2, 2)
    assert L.is_lower
    assert U.is_upper
    assert D.is_diagonal
    di, d = D.inv_den()
    assert P.matmul(A).rmul(d) == L.matmul(di).matmul(U)


def test_fflu_rank_deficient():
    A = DomainMatrix([[ZZ(1), ZZ(2)], [ZZ(2), ZZ(4)]], (2, 2), ZZ)
    P, L, D, U = A.fflu()
    assert P.shape == (2, 2)
    assert L.shape == (2, 2)
    assert D.shape == (2, 2)
    assert U.shape == (2, 2)
    assert U.getitem_sympy(1, 1) == 0

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\tests\test_polyutils.py ===
"""Tests for useful utilities for higher level polynomial classes. """

from sympy.core.mul import Mul
from sympy.core.numbers import (Integer, pi)
from sympy.core.relational import Eq
from sympy.core.singleton import S
from sympy.core.symbol import (Symbol, symbols)
from sympy.functions.elementary.exponential import exp
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.trigonometric import (cos, sin)
from sympy.integrals.integrals import Integral
from sympy.testing.pytest import raises

from sympy.polys.polyutils import (
    _nsort,
    _sort_gens,
    _unify_gens,
    _analyze_gens,
    _sort_factors,
    parallel_dict_from_expr,
    dict_from_expr,
)

from sympy.polys.polyerrors import PolynomialError

from sympy.polys.domains import ZZ

x, y, z, p, q, r, s, t, u, v, w = symbols('x,y,z,p,q,r,s,t,u,v,w')
A, B = symbols('A,B', commutative=False)


def test__nsort():
    # issue 6137
    r = S('''[3/2 + sqrt(-14/3 - 2*(-415/216 + 13*I/12)**(1/3) - 4/sqrt(-7/3 +
    61/(18*(-415/216 + 13*I/12)**(1/3)) + 2*(-415/216 + 13*I/12)**(1/3)) -
    61/(18*(-415/216 + 13*I/12)**(1/3)))/2 - sqrt(-7/3 + 61/(18*(-415/216
    + 13*I/12)**(1/3)) + 2*(-415/216 + 13*I/12)**(1/3))/2, 3/2 - sqrt(-7/3
    + 61/(18*(-415/216 + 13*I/12)**(1/3)) + 2*(-415/216 +
    13*I/12)**(1/3))/2 - sqrt(-14/3 - 2*(-415/216 + 13*I/12)**(1/3) -
    4/sqrt(-7/3 + 61/(18*(-415/216 + 13*I/12)**(1/3)) + 2*(-415/216 +
    13*I/12)**(1/3)) - 61/(18*(-415/216 + 13*I/12)**(1/3)))/2, 3/2 +
    sqrt(-14/3 - 2*(-415/216 + 13*I/12)**(1/3) + 4/sqrt(-7/3 +
    61/(18*(-415/216 + 13*I/12)**(1/3)) + 2*(-415/216 + 13*I/12)**(1/3)) -
    61/(18*(-415/216 + 13*I/12)**(1/3)))/2 + sqrt(-7/3 + 61/(18*(-415/216
    + 13*I/12)**(1/3)) + 2*(-415/216 + 13*I/12)**(1/3))/2, 3/2 + sqrt(-7/3
    + 61/(18*(-415/216 + 13*I/12)**(1/3)) + 2*(-415/216 +
    13*I/12)**(1/3))/2 - sqrt(-14/3 - 2*(-415/216 + 13*I/12)**(1/3) +
    4/sqrt(-7/3 + 61/(18*(-415/216 + 13*I/12)**(1/3)) + 2*(-415/216 +
    13*I/12)**(1/3)) - 61/(18*(-415/216 + 13*I/12)**(1/3)))/2]''')
    ans = [r[1], r[0], r[-1], r[-2]]
    assert _nsort(r) == ans
    assert len(_nsort(r, separated=True)[0]) == 0
    b, c, a = exp(-1000), exp(-999), exp(-1001)
    assert _nsort((b, c, a)) == [a, b, c]
    # issue 12560
    a = cos(1)**2 + sin(1)**2 - 1
    assert _nsort([a]) == [a]


def test__sort_gens():
    assert _sort_gens([]) == ()

    assert _sort_gens([x]) == (x,)
    assert _sort_gens([p]) == (p,)
    assert _sort_gens([q]) == (q,)

    assert _sort_gens([x, p]) == (x, p)
    assert _sort_gens([p, x]) == (x, p)
    assert _sort_gens([q, p]) == (p, q)

    assert _sort_gens([q, p, x]) == (x, p, q)

    assert _sort_gens([x, p, q], wrt=x) == (x, p, q)
    assert _sort_gens([x, p, q], wrt=p) == (p, x, q)
    assert _sort_gens([x, p, q], wrt=q) == (q, x, p)

    assert _sort_gens([x, p, q], wrt='x') == (x, p, q)
    assert _sort_gens([x, p, q], wrt='p') == (p, x, q)
    assert _sort_gens([x, p, q], wrt='q') == (q, x, p)

    assert _sort_gens([x, p, q], wrt='x,q') == (x, q, p)
    assert _sort_gens([x, p, q], wrt='q,x') == (q, x, p)
    assert _sort_gens([x, p, q], wrt='p,q') == (p, q, x)
    assert _sort_gens([x, p, q], wrt='q,p') == (q, p, x)

    assert _sort_gens([x, p, q], wrt='x, q') == (x, q, p)
    assert _sort_gens([x, p, q], wrt='q, x') == (q, x, p)
    assert _sort_gens([x, p, q], wrt='p, q') == (p, q, x)
    assert _sort_gens([x, p, q], wrt='q, p') == (q, p, x)

    assert _sort_gens([x, p, q], wrt=[x, 'q']) == (x, q, p)
    assert _sort_gens([x, p, q], wrt=[q, 'x']) == (q, x, p)
    assert _sort_gens([x, p, q], wrt=[p, 'q']) == (p, q, x)
    assert _sort_gens([x, p, q], wrt=[q, 'p']) == (q, p, x)

    assert _sort_gens([x, p, q], wrt=['x', 'q']) == (x, q, p)
    assert _sort_gens([x, p, q], wrt=['q', 'x']) == (q, x, p)
    assert _sort_gens([x, p, q], wrt=['p', 'q']) == (p, q, x)
    assert _sort_gens([x, p, q], wrt=['q', 'p']) == (q, p, x)

    assert _sort_gens([x, p, q], sort='x > p > q') == (x, p, q)
    assert _sort_gens([x, p, q], sort='p > x > q') == (p, x, q)
    assert _sort_gens([x, p, q], sort='p > q > x') == (p, q, x)

    assert _sort_gens([x, p, q], wrt='x', sort='q > p') == (x, q, p)
    assert _sort_gens([x, p, q], wrt='p', sort='q > x') == (p, q, x)
    assert _sort_gens([x, p, q], wrt='q', sort='p > x') == (q, p, x)

    # https://github.com/sympy/sympy/issues/19353
    n1 = Symbol('\n1')
    assert _sort_gens([n1]) == (n1,)
    assert _sort_gens([x, n1]) == (x, n1)

    X = symbols('x0,x1,x2,x10,x11,x12,x20,x21,x22')

    assert _sort_gens(X) == X


def test__unify_gens():
    assert _unify_gens([], []) == ()

    assert _unify_gens([x], [x]) == (x,)
    assert _unify_gens([y], [y]) == (y,)

    assert _unify_gens([x, y], [x]) == (x, y)
    assert _unify_gens([x], [x, y]) == (x, y)

    assert _unify_gens([x, y], [x, y]) == (x, y)
    assert _unify_gens([y, x], [y, x]) == (y, x)

    assert _unify_gens([x], [y]) == (x, y)
    assert _unify_gens([y], [x]) == (y, x)

    assert _unify_gens([x], [y, x]) == (y, x)
    assert _unify_gens([y, x], [x]) == (y, x)

    assert _unify_gens([x, y, z], [x, y, z]) == (x, y, z)
    assert _unify_gens([z, y, x], [x, y, z]) == (z, y, x)
    assert _unify_gens([x, y, z], [z, y, x]) == (x, y, z)
    assert _unify_gens([z, y, x], [z, y, x]) == (z, y, x)

    assert _unify_gens([x, y, z], [t, x, p, q, z]) == (t, x, y, p, q, z)


def test__analyze_gens():
    assert _analyze_gens((x, y, z)) == (x, y, z)
    assert _analyze_gens([x, y, z]) == (x, y, z)

    assert _analyze_gens(([x, y, z],)) == (x, y, z)
    assert _analyze_gens(((x, y, z),)) == (x, y, z)


def test__sort_factors():
    assert _sort_factors([], multiple=True) == []
    assert _sort_factors([], multiple=False) == []

    F = [[1, 2, 3], [1, 2], [1]]
    G = [[1], [1, 2], [1, 2, 3]]

    assert _sort_factors(F, multiple=False) == G

    F = [[1, 2], [1, 2, 3], [1, 2], [1]]
    G = [[1], [1, 2], [1, 2], [1, 2, 3]]

    assert _sort_factors(F, multiple=False) == G

    F = [[2, 2], [1, 2, 3], [1, 2], [1]]
    G = [[1], [1, 2], [2, 2], [1, 2, 3]]

    assert _sort_factors(F, multiple=False) == G

    F = [([1, 2, 3], 1), ([1, 2], 1), ([1], 1)]
    G = [([1], 1), ([1, 2], 1), ([1, 2, 3], 1)]

    assert _sort_factors(F, multiple=True) == G

    F = [([1, 2], 1), ([1, 2, 3], 1), ([1, 2], 1), ([1], 1)]
    G = [([1], 1), ([1, 2], 1), ([1, 2], 1), ([1, 2, 3], 1)]

    assert _sort_factors(F, multiple=True) == G

    F = [([2, 2], 1), ([1, 2, 3], 1), ([1, 2], 1), ([1], 1)]
    G = [([1], 1), ([1, 2], 1), ([2, 2], 1), ([1, 2, 3], 1)]

    assert _sort_factors(F, multiple=True) == G

    F = [([2, 2], 1), ([1, 2, 3], 1), ([1, 2], 2), ([1], 1)]
    G = [([1], 1), ([2, 2], 1), ([1, 2], 2), ([1, 2, 3], 1)]

    assert _sort_factors(F, multiple=True) == G


def test__dict_from_expr_if_gens():
    assert dict_from_expr(
        Integer(17), gens=(x,)) == ({(0,): Integer(17)}, (x,))
    assert dict_from_expr(
        Integer(17), gens=(x, y)) == ({(0, 0): Integer(17)}, (x, y))
    assert dict_from_expr(
        Integer(17), gens=(x, y, z)) == ({(0, 0, 0): Integer(17)}, (x, y, z))

    assert dict_from_expr(
        Integer(-17), gens=(x,)) == ({(0,): Integer(-17)}, (x,))
    assert dict_from_expr(
        Integer(-17), gens=(x, y)) == ({(0, 0): Integer(-17)}, (x, y))
    assert dict_from_expr(Integer(
        -17), gens=(x, y, z)) == ({(0, 0, 0): Integer(-17)}, (x, y, z))

    assert dict_from_expr(
        Integer(17)*x, gens=(x,)) == ({(1,): Integer(17)}, (x,))
    assert dict_from_expr(
        Integer(17)*x, gens=(x, y)) == ({(1, 0): Integer(17)}, (x, y))
    assert dict_from_expr(Integer(
        17)*x, gens=(x, y, z)) == ({(1, 0, 0): Integer(17)}, (x, y, z))

    assert dict_from_expr(
        Integer(17)*x**7, gens=(x,)) == ({(7,): Integer(17)}, (x,))
    assert dict_from_expr(
        Integer(17)*x**7*y, gens=(x, y)) == ({(7, 1): Integer(17)}, (x, y))
    assert dict_from_expr(Integer(17)*x**7*y*z**12, gens=(
        x, y, z)) == ({(7, 1, 12): Integer(17)}, (x, y, z))

    assert dict_from_expr(x + 2*y + 3*z, gens=(x,)) == \
        ({(1,): Integer(1), (0,): 2*y + 3*z}, (x,))
    assert dict_from_expr(x + 2*y + 3*z, gens=(x, y)) == \
        ({(1, 0): Integer(1), (0, 1): Integer(2), (0, 0): 3*z}, (x, y))
    assert dict_from_expr(x + 2*y + 3*z, gens=(x, y, z)) == \
        ({(1, 0, 0): Integer(
            1), (0, 1, 0): Integer(2), (0, 0, 1): Integer(3)}, (x, y, z))

    assert dict_from_expr(x*y + 2*x*z + 3*y*z, gens=(x,)) == \
        ({(1,): y + 2*z, (0,): 3*y*z}, (x,))
    assert dict_from_expr(x*y + 2*x*z + 3*y*z, gens=(x, y)) == \
        ({(1, 1): Integer(1), (1, 0): 2*z, (0, 1): 3*z}, (x, y))
    assert dict_from_expr(x*y + 2*x*z + 3*y*z, gens=(x, y, z)) == \
        ({(1, 1, 0): Integer(
            1), (1, 0, 1): Integer(2), (0, 1, 1): Integer(3)}, (x, y, z))

    assert dict_from_expr(2**y*x, gens=(x,)) == ({(1,): 2**y}, (x,))
    assert dict_from_expr(Integral(x, (x, 1, 2)) + x) == (
        {(0, 1): 1, (1, 0): 1}, (x, Integral(x, (x, 1, 2))))
    raises(PolynomialError, lambda: dict_from_expr(2**y*x, gens=(x, y)))


def test__dict_from_expr_no_gens():
    assert dict_from_expr(Integer(17)) == ({(): Integer(17)}, ())

    assert dict_from_expr(x) == ({(1,): Integer(1)}, (x,))
    assert dict_from_expr(y) == ({(1,): Integer(1)}, (y,))

    assert dict_from_expr(x*y) == ({(1, 1): Integer(1)}, (x, y))
    assert dict_from_expr(
        x + y) == ({(1, 0): Integer(1), (0, 1): Integer(1)}, (x, y))

    assert dict_from_expr(sqrt(2)) == ({(1,): Integer(1)}, (sqrt(2),))
    assert dict_from_expr(sqrt(2), greedy=False) == ({(): sqrt(2)}, ())

    assert dict_from_expr(x*y, domain=ZZ[x]) == ({(1,): x}, (y,))
    assert dict_from_expr(x*y, domain=ZZ[y]) == ({(1,): y}, (x,))

    assert dict_from_expr(3*sqrt(
        2)*pi*x*y, extension=None) == ({(1, 1, 1, 1): 3}, (x, y, pi, sqrt(2)))
    assert dict_from_expr(3*sqrt(
        2)*pi*x*y, extension=True) == ({(1, 1, 1): 3*sqrt(2)}, (x, y, pi))

    assert dict_from_expr(3*sqrt(
        2)*pi*x*y, extension=True) == ({(1, 1, 1): 3*sqrt(2)}, (x, y, pi))

    f = cos(x)*sin(x) + cos(x)*sin(y) + cos(y)*sin(x) + cos(y)*sin(y)

    assert dict_from_expr(f) == ({(0, 1, 0, 1): 1, (0, 1, 1, 0): 1,
        (1, 0, 0, 1): 1, (1, 0, 1, 0): 1}, (cos(x), cos(y), sin(x), sin(y)))


def test__parallel_dict_from_expr_if_gens():
    assert parallel_dict_from_expr([x + 2*y + 3*z, Integer(7)], gens=(x,)) == \
        ([{(1,): Integer(1), (0,): 2*y + 3*z}, {(0,): Integer(7)}], (x,))


def test__parallel_dict_from_expr_no_gens():
    assert parallel_dict_from_expr([x*y, Integer(3)]) == \
        ([{(1, 1): Integer(1)}, {(0, 0): Integer(3)}], (x, y))
    assert parallel_dict_from_expr([x*y, 2*z, Integer(3)]) == \
        ([{(1, 1, 0): Integer(
            1)}, {(0, 0, 1): Integer(2)}, {(0, 0, 0): Integer(3)}], (x, y, z))
    assert parallel_dict_from_expr((Mul(x, x**2, evaluate=False),)) == \
        ([{(3,): 1}], (x,))


def test_parallel_dict_from_expr():
    assert parallel_dict_from_expr([Eq(x, 1), Eq(
        x**2, 2)]) == ([{(0,): -Integer(1), (1,): Integer(1)},
                        {(0,): -Integer(2), (2,): Integer(1)}], (x,))
    raises(PolynomialError, lambda: parallel_dict_from_expr([A*B - B*A]))


def test_dict_from_expr():
    assert dict_from_expr(Eq(x, 1)) == \
        ({(0,): -Integer(1), (1,): Integer(1)}, (x,))
    raises(PolynomialError, lambda: dict_from_expr(A*B - B*A))
    raises(PolynomialError, lambda: dict_from_expr(S.true))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\libs\test_lib.py ===
import pickle

import numpy as np
import pytest

from pandas._libs import (
    Timedelta,
    lib,
    writers as libwriters,
)
from pandas.compat import IS64

from pandas import Index
import pandas._testing as tm


class TestMisc:
    def test_max_len_string_array(self):
        arr = a = np.array(["foo", "b", np.nan], dtype="object")
        assert libwriters.max_len_string_array(arr) == 3

        # unicode
        arr = a.astype("U").astype(object)
        assert libwriters.max_len_string_array(arr) == 3

        # bytes for python3
        arr = a.astype("S").astype(object)
        assert libwriters.max_len_string_array(arr) == 3

        # raises
        msg = "No matching signature found"
        with pytest.raises(TypeError, match=msg):
            libwriters.max_len_string_array(arr.astype("U"))

    def test_fast_unique_multiple_list_gen_sort(self):
        keys = [["p", "a"], ["n", "d"], ["a", "s"]]

        gen = (key for key in keys)
        expected = np.array(["a", "d", "n", "p", "s"])
        out = lib.fast_unique_multiple_list_gen(gen, sort=True)
        tm.assert_numpy_array_equal(np.array(out), expected)

        gen = (key for key in keys)
        expected = np.array(["p", "a", "n", "d", "s"])
        out = lib.fast_unique_multiple_list_gen(gen, sort=False)
        tm.assert_numpy_array_equal(np.array(out), expected)

    def test_fast_multiget_timedelta_resos(self):
        # This will become relevant for test_constructor_dict_timedelta64_index
        #  once Timedelta constructor preserves reso when passed a
        #  np.timedelta64 object
        td = Timedelta(days=1)

        mapping1 = {td: 1}
        mapping2 = {td.as_unit("s"): 1}

        oindex = Index([td * n for n in range(3)])._values.astype(object)

        expected = lib.fast_multiget(mapping1, oindex)
        result = lib.fast_multiget(mapping2, oindex)
        tm.assert_numpy_array_equal(result, expected)

        # case that can't be cast to td64ns
        td = Timedelta(np.timedelta64(146000, "D"))
        assert hash(td) == hash(td.as_unit("ms"))
        assert hash(td) == hash(td.as_unit("us"))
        mapping1 = {td: 1}
        mapping2 = {td.as_unit("ms"): 1}

        oindex = Index([td * n for n in range(3)])._values.astype(object)

        expected = lib.fast_multiget(mapping1, oindex)
        result = lib.fast_multiget(mapping2, oindex)
        tm.assert_numpy_array_equal(result, expected)


class TestIndexing:
    def test_maybe_indices_to_slice_left_edge(self):
        target = np.arange(100)

        # slice
        indices = np.array([], dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

    @pytest.mark.parametrize("end", [1, 2, 5, 20, 99])
    @pytest.mark.parametrize("step", [1, 2, 4])
    def test_maybe_indices_to_slice_left_edge_not_slice_end_steps(self, end, step):
        target = np.arange(100)
        indices = np.arange(0, end, step, dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

        # reverse
        indices = indices[::-1]
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

    @pytest.mark.parametrize(
        "case", [[2, 1, 2, 0], [2, 2, 1, 0], [0, 1, 2, 1], [-2, 0, 2], [2, 0, -2]]
    )
    def test_maybe_indices_to_slice_left_edge_not_slice(self, case):
        # not slice
        target = np.arange(100)
        indices = np.array(case, dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert not isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(maybe_slice, indices)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

    @pytest.mark.parametrize("start", [0, 2, 5, 20, 97, 98])
    @pytest.mark.parametrize("step", [1, 2, 4])
    def test_maybe_indices_to_slice_right_edge(self, start, step):
        target = np.arange(100)

        # slice
        indices = np.arange(start, 99, step, dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

        # reverse
        indices = indices[::-1]
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

    def test_maybe_indices_to_slice_right_edge_not_slice(self):
        # not slice
        target = np.arange(100)
        indices = np.array([97, 98, 99, 100], dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert not isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(maybe_slice, indices)

        msg = "index 100 is out of bounds for axis (0|1) with size 100"

        with pytest.raises(IndexError, match=msg):
            target[indices]
        with pytest.raises(IndexError, match=msg):
            target[maybe_slice]

        indices = np.array([100, 99, 98, 97], dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert not isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(maybe_slice, indices)

        with pytest.raises(IndexError, match=msg):
            target[indices]
        with pytest.raises(IndexError, match=msg):
            target[maybe_slice]

    @pytest.mark.parametrize(
        "case", [[99, 97, 99, 96], [99, 99, 98, 97], [98, 98, 97, 96]]
    )
    def test_maybe_indices_to_slice_right_edge_cases(self, case):
        target = np.arange(100)
        indices = np.array(case, dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert not isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(maybe_slice, indices)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

    @pytest.mark.parametrize("step", [1, 2, 4, 5, 8, 9])
    def test_maybe_indices_to_slice_both_edges(self, step):
        target = np.arange(10)

        # slice
        indices = np.arange(0, 9, step, dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))
        assert isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

        # reverse
        indices = indices[::-1]
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))
        assert isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

    @pytest.mark.parametrize("case", [[4, 2, 0, -2], [2, 2, 1, 0], [0, 1, 2, 1]])
    def test_maybe_indices_to_slice_both_edges_not_slice(self, case):
        # not slice
        target = np.arange(10)
        indices = np.array(case, dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))
        assert not isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(maybe_slice, indices)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

    @pytest.mark.parametrize("start, end", [(2, 10), (5, 25), (65, 97)])
    @pytest.mark.parametrize("step", [1, 2, 4, 20])
    def test_maybe_indices_to_slice_middle(self, start, end, step):
        target = np.arange(100)

        # slice
        indices = np.arange(start, end, step, dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

        # reverse
        indices = indices[::-1]
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

    @pytest.mark.parametrize(
        "case", [[14, 12, 10, 12], [12, 12, 11, 10], [10, 11, 12, 11]]
    )
    def test_maybe_indices_to_slice_middle_not_slice(self, case):
        # not slice
        target = np.arange(100)
        indices = np.array(case, dtype=np.intp)
        maybe_slice = lib.maybe_indices_to_slice(indices, len(target))

        assert not isinstance(maybe_slice, slice)
        tm.assert_numpy_array_equal(maybe_slice, indices)
        tm.assert_numpy_array_equal(target[indices], target[maybe_slice])

    def test_maybe_booleans_to_slice(self):
        arr = np.array([0, 0, 1, 1, 1, 0, 1], dtype=np.uint8)
        result = lib.maybe_booleans_to_slice(arr)
        assert result.dtype == np.bool_

        result = lib.maybe_booleans_to_slice(arr[:0])
        assert result == slice(0, 0)

    def test_get_reverse_indexer(self):
        indexer = np.array([-1, -1, 1, 2, 0, -1, 3, 4], dtype=np.intp)
        result = lib.get_reverse_indexer(indexer, 5)
        expected = np.array([4, 2, 3, 6, 7], dtype=np.intp)
        tm.assert_numpy_array_equal(result, expected)

    @pytest.mark.parametrize("dtype", ["int64", "int32"])
    def test_is_range_indexer(self, dtype):
        # GH#50592
        left = np.arange(0, 100, dtype=dtype)
        assert lib.is_range_indexer(left, 100)

    @pytest.mark.skipif(
        not IS64,
        reason="2**31 is too big for Py_ssize_t on 32-bit. "
        "It doesn't matter though since you cannot create an array that long on 32-bit",
    )
    @pytest.mark.parametrize("dtype", ["int64", "int32"])
    def test_is_range_indexer_big_n(self, dtype):
        # GH53616
        left = np.arange(0, 100, dtype=dtype)

        assert not lib.is_range_indexer(left, 2**31)

    @pytest.mark.parametrize("dtype", ["int64", "int32"])
    def test_is_range_indexer_not_equal(self, dtype):
        # GH#50592
        left = np.array([1, 2], dtype=dtype)
        assert not lib.is_range_indexer(left, 2)

    @pytest.mark.parametrize("dtype", ["int64", "int32"])
    def test_is_range_indexer_not_equal_shape(self, dtype):
        # GH#50592
        left = np.array([0, 1, 2], dtype=dtype)
        assert not lib.is_range_indexer(left, 2)


def test_cache_readonly_preserve_docstrings():
    # GH18197
    assert Index.hasnans.__doc__ is not None


def test_no_default_pickle():
    # GH#40397
    obj = tm.round_trip_pickle(lib.no_default)
    assert obj is lib.no_default


def test_ensure_string_array_copy():
    # ensure the original array is not modified in case of copy=False with
    # pickle-roundtripped object dtype array
    # https://github.com/pandas-dev/pandas/issues/54654
    arr = np.array(["a", None], dtype=object)
    arr = pickle.loads(pickle.dumps(arr))
    result = lib.ensure_string_array(arr, copy=False)
    assert not np.shares_memory(arr, result)
    assert arr[1] is None
    assert result[1] is np.nan

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\reshape\concat\test_empty.py ===
import numpy as np
import pytest

from pandas._config import using_string_dtype

import pandas as pd
from pandas import (
    DataFrame,
    RangeIndex,
    Series,
    concat,
    date_range,
)
import pandas._testing as tm


class TestEmptyConcat:
    def test_handle_empty_objects(self, sort, using_infer_string):
        df = DataFrame(
            np.random.default_rng(2).standard_normal((10, 4)), columns=list("abcd")
        )

        dfcopy = df[:5].copy()
        dfcopy["foo"] = "bar"
        empty = df[5:5]

        frames = [dfcopy, empty, empty, df[5:]]
        concatted = concat(frames, axis=0, sort=sort)

        expected = df.reindex(columns=["a", "b", "c", "d", "foo"])
        expected["foo"] = expected["foo"].astype(
            object if not using_infer_string else "str"
        )
        expected.loc[0:4, "foo"] = "bar"

        tm.assert_frame_equal(concatted, expected)

        # empty as first element with time series
        # GH3259
        df = DataFrame(
            {"A": range(10000)}, index=date_range("20130101", periods=10000, freq="s")
        )
        empty = DataFrame()
        result = concat([df, empty], axis=1)
        tm.assert_frame_equal(result, df)
        result = concat([empty, df], axis=1)
        tm.assert_frame_equal(result, df)

        result = concat([df, empty])
        tm.assert_frame_equal(result, df)
        result = concat([empty, df])
        tm.assert_frame_equal(result, df)

    def test_concat_empty_series(self):
        # GH 11082
        s1 = Series([1, 2, 3], name="x")
        s2 = Series(name="y", dtype="float64")
        res = concat([s1, s2], axis=1)
        exp = DataFrame(
            {"x": [1, 2, 3], "y": [np.nan, np.nan, np.nan]},
            index=RangeIndex(3),
        )
        tm.assert_frame_equal(res, exp)

        s1 = Series([1, 2, 3], name="x")
        s2 = Series(name="y", dtype="float64")
        msg = "The behavior of array concatenation with empty entries is deprecated"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            res = concat([s1, s2], axis=0)
        # name will be reset
        exp = Series([1, 2, 3])
        tm.assert_series_equal(res, exp)

        # empty Series with no name
        s1 = Series([1, 2, 3], name="x")
        s2 = Series(name=None, dtype="float64")
        res = concat([s1, s2], axis=1)
        exp = DataFrame(
            {"x": [1, 2, 3], 0: [np.nan, np.nan, np.nan]},
            columns=["x", 0],
            index=RangeIndex(3),
        )
        tm.assert_frame_equal(res, exp)

    @pytest.mark.parametrize("tz", [None, "UTC"])
    @pytest.mark.parametrize("values", [[], [1, 2, 3]])
    def test_concat_empty_series_timelike(self, tz, values):
        # GH 18447

        first = Series([], dtype="M8[ns]").dt.tz_localize(tz)
        dtype = None if values else np.float64
        second = Series(values, dtype=dtype)

        expected = DataFrame(
            {
                0: Series([pd.NaT] * len(values), dtype="M8[ns]").dt.tz_localize(tz),
                1: values,
            }
        )
        result = concat([first, second], axis=1)
        tm.assert_frame_equal(result, expected)

    @pytest.mark.parametrize(
        "left,right,expected",
        [
            # booleans
            (np.bool_, np.int32, np.object_),  # changed from int32 in 2.0 GH#39817
            (np.bool_, np.float32, np.object_),
            # datetime-like
            ("m8[ns]", np.bool_, np.object_),
            ("m8[ns]", np.int64, np.object_),
            ("M8[ns]", np.bool_, np.object_),
            ("M8[ns]", np.int64, np.object_),
            # categorical
            ("category", "category", "category"),
            ("category", "object", "object"),
        ],
    )
    def test_concat_empty_series_dtypes(self, left, right, expected):
        # GH#39817, GH#45101
        result = concat([Series(dtype=left), Series(dtype=right)])
        assert result.dtype == expected

    @pytest.mark.parametrize(
        "dtype", ["float64", "int8", "uint8", "bool", "m8[ns]", "M8[ns]"]
    )
    def test_concat_empty_series_dtypes_match_roundtrips(self, dtype):
        dtype = np.dtype(dtype)

        result = concat([Series(dtype=dtype)])
        assert result.dtype == dtype

        result = concat([Series(dtype=dtype), Series(dtype=dtype)])
        assert result.dtype == dtype

    @pytest.mark.parametrize("dtype", ["float64", "int8", "uint8", "m8[ns]", "M8[ns]"])
    @pytest.mark.parametrize(
        "dtype2",
        ["float64", "int8", "uint8", "m8[ns]", "M8[ns]"],
    )
    def test_concat_empty_series_dtypes_roundtrips(self, dtype, dtype2):
        # round-tripping with self & like self
        if dtype == dtype2:
            pytest.skip("same dtype is not applicable for test")

        def int_result_type(dtype, dtype2):
            typs = {dtype.kind, dtype2.kind}
            if not len(typs - {"i", "u", "b"}) and (
                dtype.kind == "i" or dtype2.kind == "i"
            ):
                return "i"
            elif not len(typs - {"u", "b"}) and (
                dtype.kind == "u" or dtype2.kind == "u"
            ):
                return "u"
            return None

        def float_result_type(dtype, dtype2):
            typs = {dtype.kind, dtype2.kind}
            if not len(typs - {"f", "i", "u"}) and (
                dtype.kind == "f" or dtype2.kind == "f"
            ):
                return "f"
            return None

        def get_result_type(dtype, dtype2):
            result = float_result_type(dtype, dtype2)
            if result is not None:
                return result
            result = int_result_type(dtype, dtype2)
            if result is not None:
                return result
            return "O"

        dtype = np.dtype(dtype)
        dtype2 = np.dtype(dtype2)
        expected = get_result_type(dtype, dtype2)
        result = concat([Series(dtype=dtype), Series(dtype=dtype2)]).dtype
        assert result.kind == expected

    def test_concat_empty_series_dtypes_triple(self):
        assert (
            concat(
                [Series(dtype="M8[ns]"), Series(dtype=np.bool_), Series(dtype=np.int64)]
            ).dtype
            == np.object_
        )

    def test_concat_empty_series_dtype_category_with_array(self):
        # GH#18515
        assert (
            concat(
                [Series(np.array([]), dtype="category"), Series(dtype="float64")]
            ).dtype
            == "float64"
        )

    def test_concat_empty_series_dtypes_sparse(self):
        result = concat(
            [
                Series(dtype="float64").astype("Sparse"),
                Series(dtype="float64").astype("Sparse"),
            ]
        )
        assert result.dtype == "Sparse[float64]"

        result = concat(
            [Series(dtype="float64").astype("Sparse"), Series(dtype="float64")]
        )
        expected = pd.SparseDtype(np.float64)
        assert result.dtype == expected

        result = concat(
            [Series(dtype="float64").astype("Sparse"), Series(dtype="object")]
        )
        expected = pd.SparseDtype("object")
        assert result.dtype == expected

    def test_concat_empty_df_object_dtype(self):
        # GH 9149
        df_1 = DataFrame({"Row": [0, 1, 1], "EmptyCol": np.nan, "NumberCol": [1, 2, 3]})
        df_2 = DataFrame(columns=df_1.columns)
        result = concat([df_1, df_2], axis=0)
        expected = df_1.astype(object)
        tm.assert_frame_equal(result, expected)

    def test_concat_empty_dataframe_dtypes(self):
        df = DataFrame(columns=list("abc"))
        df["a"] = df["a"].astype(np.bool_)
        df["b"] = df["b"].astype(np.int32)
        df["c"] = df["c"].astype(np.float64)

        result = concat([df, df])
        assert result["a"].dtype == np.bool_
        assert result["b"].dtype == np.int32
        assert result["c"].dtype == np.float64

        result = concat([df, df.astype(np.float64)])
        assert result["a"].dtype == np.object_
        assert result["b"].dtype == np.float64
        assert result["c"].dtype == np.float64

    # triggers warning about empty entries
    @pytest.mark.xfail(using_string_dtype(), reason="TODO(infer_string)")
    def test_concat_inner_join_empty(self):
        # GH 15328
        df_empty = DataFrame()
        df_a = DataFrame({"a": [1, 2]}, index=[0, 1], dtype="int64")
        df_expected = DataFrame({"a": []}, index=RangeIndex(0), dtype="int64")

        result = concat([df_a, df_empty], axis=1, join="inner")
        tm.assert_frame_equal(result, df_expected)

        result = concat([df_a, df_empty], axis=1, join="outer")
        tm.assert_frame_equal(result, df_a)

    def test_empty_dtype_coerce(self):
        # xref to #12411
        # xref to #12045
        # xref to #11594
        # see below

        # 10571
        df1 = DataFrame(data=[[1, None], [2, None]], columns=["a", "b"])
        df2 = DataFrame(data=[[3, None], [4, None]], columns=["a", "b"])
        result = concat([df1, df2])
        expected = df1.dtypes
        tm.assert_series_equal(result.dtypes, expected)

    def test_concat_empty_dataframe(self):
        # 39037
        df1 = DataFrame(columns=["a", "b"])
        df2 = DataFrame(columns=["b", "c"])
        result = concat([df1, df2, df1])
        expected = DataFrame(columns=["a", "b", "c"])
        tm.assert_frame_equal(result, expected)

        df3 = DataFrame(columns=["a", "b"])
        df4 = DataFrame(columns=["b"])
        result = concat([df3, df4])
        expected = DataFrame(columns=["a", "b"])
        tm.assert_frame_equal(result, expected)

    def test_concat_empty_dataframe_different_dtypes(self, using_infer_string):
        # 39037
        df1 = DataFrame({"a": [1, 2, 3], "b": ["a", "b", "c"]})
        df2 = DataFrame({"a": [1, 2, 3]})

        result = concat([df1[:0], df2[:0]])
        assert result["a"].dtype == np.int64
        assert result["b"].dtype == np.object_ if not using_infer_string else "str"

    def test_concat_to_empty_ea(self):
        """48510 `concat` to an empty EA should maintain type EA dtype."""
        df_empty = DataFrame({"a": pd.array([], dtype=pd.Int64Dtype())})
        df_new = DataFrame({"a": pd.array([1, 2, 3], dtype=pd.Int64Dtype())})
        expected = df_new.copy()
        result = concat([df_empty, df_new])
        tm.assert_frame_equal(result, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\series\test_api.py ===
import inspect
import pydoc

import numpy as np
import pytest

import pandas as pd
from pandas import (
    DataFrame,
    Index,
    Series,
    date_range,
    period_range,
    timedelta_range,
)
import pandas._testing as tm


class TestSeriesMisc:
    def test_tab_completion(self):
        # GH 9910
        s = Series(list("abcd"))
        # Series of str values should have .str but not .dt/.cat in __dir__
        assert "str" in dir(s)
        assert "dt" not in dir(s)
        assert "cat" not in dir(s)

    def test_tab_completion_dt(self):
        # similarly for .dt
        s = Series(date_range("1/1/2015", periods=5))
        assert "dt" in dir(s)
        assert "str" not in dir(s)
        assert "cat" not in dir(s)

    def test_tab_completion_cat(self):
        # Similarly for .cat, but with the twist that str and dt should be
        # there if the categories are of that type first cat and str.
        s = Series(list("abbcd"), dtype="category")
        assert "cat" in dir(s)
        assert "str" in dir(s)  # as it is a string categorical
        assert "dt" not in dir(s)

    def test_tab_completion_cat_str(self):
        # similar to cat and str
        s = Series(date_range("1/1/2015", periods=5)).astype("category")
        assert "cat" in dir(s)
        assert "str" not in dir(s)
        assert "dt" in dir(s)  # as it is a datetime categorical

    def test_tab_completion_with_categorical(self):
        # test the tab completion display
        ok_for_cat = [
            "categories",
            "codes",
            "ordered",
            "set_categories",
            "add_categories",
            "remove_categories",
            "rename_categories",
            "reorder_categories",
            "remove_unused_categories",
            "as_ordered",
            "as_unordered",
        ]

        s = Series(list("aabbcde")).astype("category")
        results = sorted({r for r in s.cat.__dir__() if not r.startswith("_")})
        tm.assert_almost_equal(results, sorted(set(ok_for_cat)))

    @pytest.mark.parametrize(
        "index",
        [
            Index(list("ab") * 5, dtype="category"),
            Index([str(i) for i in range(10)]),
            Index(["foo", "bar", "baz"] * 2),
            date_range("2020-01-01", periods=10),
            period_range("2020-01-01", periods=10, freq="D"),
            timedelta_range("1 day", periods=10),
            Index(np.arange(10), dtype=np.uint64),
            Index(np.arange(10), dtype=np.int64),
            Index(np.arange(10), dtype=np.float64),
            Index([True, False]),
            Index([f"a{i}" for i in range(101)]),
            pd.MultiIndex.from_tuples(zip("ABCD", "EFGH")),
            pd.MultiIndex.from_tuples(zip([0, 1, 2, 3], "EFGH")),
        ],
    )
    def test_index_tab_completion(self, index):
        # dir contains string-like values of the Index.
        s = Series(index=index, dtype=object)
        dir_s = dir(s)
        for i, x in enumerate(s.index.unique(level=0)):
            if i < 100:
                assert not isinstance(x, str) or not x.isidentifier() or x in dir_s
            else:
                assert x not in dir_s

    @pytest.mark.parametrize("ser", [Series(dtype=object), Series([1])])
    def test_not_hashable(self, ser):
        msg = "unhashable type: 'Series'"
        with pytest.raises(TypeError, match=msg):
            hash(ser)

    def test_contains(self, datetime_series):
        tm.assert_contains_all(datetime_series.index, datetime_series)

    def test_axis_alias(self):
        s = Series([1, 2, np.nan])
        tm.assert_series_equal(s.dropna(axis="rows"), s.dropna(axis="index"))
        assert s.dropna().sum("rows") == 3
        assert s._get_axis_number("rows") == 0
        assert s._get_axis_name("rows") == "index"

    def test_class_axis(self):
        # https://github.com/pandas-dev/pandas/issues/18147
        # no exception and no empty docstring
        assert pydoc.getdoc(Series.index)

    def test_ndarray_compat(self):
        # test numpy compat with Series as sub-class of NDFrame
        tsdf = DataFrame(
            np.random.default_rng(2).standard_normal((1000, 3)),
            columns=["A", "B", "C"],
            index=date_range("1/1/2000", periods=1000),
        )

        def f(x):
            return x[x.idxmax()]

        result = tsdf.apply(f)
        expected = tsdf.max()
        tm.assert_series_equal(result, expected)

    def test_ndarray_compat_like_func(self):
        # using an ndarray like function
        s = Series(np.random.default_rng(2).standard_normal(10))
        result = Series(np.ones_like(s))
        expected = Series(1, index=range(10), dtype="float64")
        tm.assert_series_equal(result, expected)

    def test_ndarray_compat_ravel(self):
        # ravel
        s = Series(np.random.default_rng(2).standard_normal(10))
        with tm.assert_produces_warning(FutureWarning, match="ravel is deprecated"):
            result = s.ravel(order="F")
        tm.assert_almost_equal(result, s.values.ravel(order="F"))

    def test_empty_method(self):
        s_empty = Series(dtype=object)
        assert s_empty.empty

    @pytest.mark.parametrize("dtype", ["int64", object])
    def test_empty_method_full_series(self, dtype):
        full_series = Series(index=[1], dtype=dtype)
        assert not full_series.empty

    @pytest.mark.parametrize("dtype", [None, "Int64"])
    def test_integer_series_size(self, dtype):
        # GH 25580
        s = Series(range(9), dtype=dtype)
        assert s.size == 9

    def test_attrs(self):
        s = Series([0, 1], name="abc")
        assert s.attrs == {}
        s.attrs["version"] = 1
        result = s + 1
        assert result.attrs == {"version": 1}

    def test_inspect_getmembers(self):
        # GH38782
        ser = Series(dtype=object)
        msg = "Series._data is deprecated"
        with tm.assert_produces_warning(
            DeprecationWarning, match=msg, check_stacklevel=False
        ):
            inspect.getmembers(ser)

    def test_unknown_attribute(self):
        # GH#9680
        tdi = timedelta_range(start=0, periods=10, freq="1s")
        ser = Series(np.random.default_rng(2).normal(size=10), index=tdi)
        assert "foo" not in ser.__dict__
        msg = "'Series' object has no attribute 'foo'"
        with pytest.raises(AttributeError, match=msg):
            ser.foo

    @pytest.mark.parametrize("op", ["year", "day", "second", "weekday"])
    def test_datetime_series_no_datelike_attrs(self, op, datetime_series):
        # GH#7206
        msg = f"'Series' object has no attribute '{op}'"
        with pytest.raises(AttributeError, match=msg):
            getattr(datetime_series, op)

    def test_series_datetimelike_attribute_access(self):
        # attribute access should still work!
        ser = Series({"year": 2000, "month": 1, "day": 10})
        assert ser.year == 2000
        assert ser.month == 1
        assert ser.day == 10

    def test_series_datetimelike_attribute_access_invalid(self):
        ser = Series({"year": 2000, "month": 1, "day": 10})
        msg = "'Series' object has no attribute 'weekday'"
        with pytest.raises(AttributeError, match=msg):
            ser.weekday

    @pytest.mark.filterwarnings("ignore:Downcasting object dtype arrays:FutureWarning")
    @pytest.mark.parametrize(
        "kernel, has_numeric_only",
        [
            ("skew", True),
            ("var", True),
            ("all", False),
            ("prod", True),
            ("any", False),
            ("idxmin", False),
            ("quantile", False),
            ("idxmax", False),
            ("min", True),
            ("sem", True),
            ("mean", True),
            ("nunique", False),
            ("max", True),
            ("sum", True),
            ("count", False),
            ("median", True),
            ("std", True),
            ("backfill", False),
            ("rank", True),
            ("pct_change", False),
            ("cummax", False),
            ("shift", False),
            ("diff", False),
            ("cumsum", False),
            ("cummin", False),
            ("cumprod", False),
            ("fillna", False),
            ("ffill", False),
            ("pad", False),
            ("bfill", False),
            ("sample", False),
            ("tail", False),
            ("take", False),
            ("head", False),
            ("cov", False),
            ("corr", False),
        ],
    )
    @pytest.mark.parametrize("dtype", [bool, int, float, object])
    def test_numeric_only(self, kernel, has_numeric_only, dtype):
        # GH#47500
        ser = Series([0, 1, 1], dtype=dtype)
        if kernel == "corrwith":
            args = (ser,)
        elif kernel == "corr":
            args = (ser,)
        elif kernel == "cov":
            args = (ser,)
        elif kernel == "nth":
            args = (0,)
        elif kernel == "fillna":
            args = (True,)
        elif kernel == "fillna":
            args = ("ffill",)
        elif kernel == "take":
            args = ([0],)
        elif kernel == "quantile":
            args = (0.5,)
        else:
            args = ()
        method = getattr(ser, kernel)
        if not has_numeric_only:
            msg = (
                "(got an unexpected keyword argument 'numeric_only'"
                "|too many arguments passed in)"
            )
            with pytest.raises(TypeError, match=msg):
                method(*args, numeric_only=True)
        elif dtype is object:
            msg = f"Series.{kernel} does not allow numeric_only=True with non-numeric"
            with pytest.raises(TypeError, match=msg):
                method(*args, numeric_only=True)
        else:
            result = method(*args, numeric_only=True)
            expected = method(*args, numeric_only=False)
            if isinstance(expected, Series):
                # transformer
                tm.assert_series_equal(result, expected)
            else:
                # reducer
                assert result == expected


@pytest.mark.parametrize("converter", [int, float, complex])
def test_float_int_deprecated(converter):
    # GH 51101
    with tm.assert_produces_warning(FutureWarning):
        assert converter(Series([1])) == converter(1)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\expressions\tests\test_indexing.py ===
from sympy.concrete.summations import Sum
from sympy.core.symbol import symbols, Symbol, Dummy
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.special.tensor_functions import KroneckerDelta
from sympy.matrices.dense import eye
from sympy.matrices.expressions.blockmatrix import BlockMatrix
from sympy.matrices.expressions.hadamard import HadamardPower
from sympy.matrices.expressions.matexpr import (MatrixSymbol,
    MatrixExpr, MatrixElement)
from sympy.matrices.expressions.matpow import MatPow
from sympy.matrices.expressions.special import (ZeroMatrix, Identity,
    OneMatrix)
from sympy.matrices.expressions.trace import Trace, trace
from sympy.matrices.immutable import ImmutableMatrix
from sympy.tensor.array.expressions.array_expressions import ArrayTensorProduct
from sympy.testing.pytest import XFAIL, raises

k, l, m, n = symbols('k l m n', integer=True)
i, j = symbols('i j', integer=True)

W = MatrixSymbol('W', k, l)
X = MatrixSymbol('X', l, m)
Y = MatrixSymbol('Y', l, m)
Z = MatrixSymbol('Z', m, n)

X1 = MatrixSymbol('X1', m, m)
X2 = MatrixSymbol('X2', m, m)
X3 = MatrixSymbol('X3', m, m)
X4 = MatrixSymbol('X4', m, m)

A = MatrixSymbol('A', 2, 2)
B = MatrixSymbol('B', 2, 2)
x = MatrixSymbol('x', 1, 2)
y = MatrixSymbol('x', 2, 1)


def test_symbolic_indexing():
    x12 = X[1, 2]
    assert all(s in str(x12) for s in ['1', '2', X.name])
    # We don't care about the exact form of this.  We do want to make sure
    # that all of these features are present


def test_add_index():
    assert (X + Y)[i, j] == X[i, j] + Y[i, j]


def test_mul_index():
    assert (A*y)[0, 0] == A[0, 0]*y[0, 0] + A[0, 1]*y[1, 0]
    assert (A*B).as_mutable() == (A.as_mutable() * B.as_mutable())
    X = MatrixSymbol('X', n, m)
    Y = MatrixSymbol('Y', m, k)

    result = (X*Y)[4,2]
    expected = Sum(X[4, i]*Y[i, 2], (i, 0, m - 1))
    assert result.args[0].dummy_eq(expected.args[0], i)
    assert result.args[1][1:] == expected.args[1][1:]


def test_pow_index():
    Q = MatPow(A, 2)
    assert Q[0, 0] == A[0, 0]**2 + A[0, 1]*A[1, 0]
    n = symbols("n")
    Q2 = A**n
    assert Q2[0, 0] == 2*(
            -sqrt((A[0, 0] + A[1, 1])**2 - 4*A[0, 0]*A[1, 1] +
            4*A[0, 1]*A[1, 0])/2 + A[0, 0]/2 + A[1, 1]/2
        )**n * \
        A[0, 1]*A[1, 0]/(
            (sqrt(A[0, 0]**2 - 2*A[0, 0]*A[1, 1] + 4*A[0, 1]*A[1, 0] +
                  A[1, 1]**2) + A[0, 0] - A[1, 1])*
            sqrt(A[0, 0]**2 - 2*A[0, 0]*A[1, 1] + 4*A[0, 1]*A[1, 0] + A[1, 1]**2)
        ) - 2*(
            sqrt((A[0, 0] + A[1, 1])**2 - 4*A[0, 0]*A[1, 1] +
            4*A[0, 1]*A[1, 0])/2 + A[0, 0]/2 + A[1, 1]/2
        )**n * A[0, 1]*A[1, 0]/(
            (-sqrt(A[0, 0]**2 - 2*A[0, 0]*A[1, 1] + 4*A[0, 1]*A[1, 0] +
            A[1, 1]**2) + A[0, 0] - A[1, 1])*
            sqrt(A[0, 0]**2 - 2*A[0, 0]*A[1, 1] + 4*A[0, 1]*A[1, 0] + A[1, 1]**2)
        )


def test_transpose_index():
    assert X.T[i, j] == X[j, i]


def test_Identity_index():
    I = Identity(3)
    assert I[0, 0] == I[1, 1] == I[2, 2] == 1
    assert I[1, 0] == I[0, 1] == I[2, 1] == 0
    assert I[i, 0].delta_range == (0, 2)
    raises(IndexError, lambda: I[3, 3])


def test_block_index():
    I = Identity(3)
    Z = ZeroMatrix(3, 3)
    B = BlockMatrix([[I, I], [I, I]])
    e3 = ImmutableMatrix(eye(3))
    BB = BlockMatrix([[e3, e3], [e3, e3]])
    assert B[0, 0] == B[3, 0] == B[0, 3] == B[3, 3] == 1
    assert B[4, 3] == B[5, 1] == 0

    BB = BlockMatrix([[e3, e3], [e3, e3]])
    assert B.as_explicit() == BB.as_explicit()

    BI = BlockMatrix([[I, Z], [Z, I]])

    assert BI.as_explicit().equals(eye(6))


def test_block_index_symbolic():
    # Note that these matrices may be zero-sized and indices may be negative, which causes
    # all naive simplifications given in the comments to be invalid
    A1 = MatrixSymbol('A1', n, k)
    A2 = MatrixSymbol('A2', n, l)
    A3 = MatrixSymbol('A3', m, k)
    A4 = MatrixSymbol('A4', m, l)
    A = BlockMatrix([[A1, A2], [A3, A4]])
    assert A[0, 0] == MatrixElement(A, 0, 0)  # Cannot be A1[0, 0]
    assert A[n - 1, k - 1] == A1[n - 1, k - 1]
    assert A[n, k] == A4[0, 0]
    assert A[n + m - 1, 0] == MatrixElement(A, n + m - 1, 0)  # Cannot be A3[m - 1, 0]
    assert A[0, k + l - 1] == MatrixElement(A, 0, k + l - 1)  # Cannot be A2[0, l - 1]
    assert A[n + m - 1, k + l - 1] == MatrixElement(A, n + m - 1, k + l - 1)  # Cannot be A4[m - 1, l - 1]
    assert A[i, j] == MatrixElement(A, i, j)
    assert A[n + i, k + j] == MatrixElement(A, n + i, k + j)  # Cannot be A4[i, j]
    assert A[n - i - 1, k - j - 1] == MatrixElement(A, n - i - 1, k - j - 1)  # Cannot be A1[n - i - 1, k - j - 1]


def test_block_index_symbolic_nonzero():
    # All invalid simplifications from test_block_index_symbolic() that become valid if all
    # matrices have nonzero size and all indices are nonnegative
    k, l, m, n = symbols('k l m n', integer=True, positive=True)
    i, j = symbols('i j', integer=True, nonnegative=True)
    A1 = MatrixSymbol('A1', n, k)
    A2 = MatrixSymbol('A2', n, l)
    A3 = MatrixSymbol('A3', m, k)
    A4 = MatrixSymbol('A4', m, l)
    A = BlockMatrix([[A1, A2], [A3, A4]])
    assert A[0, 0] == A1[0, 0]
    assert A[n + m - 1, 0] == A3[m - 1, 0]
    assert A[0, k + l - 1] == A2[0, l - 1]
    assert A[n + m - 1, k + l - 1] == A4[m - 1, l - 1]
    assert A[i, j] == MatrixElement(A, i, j)
    assert A[n + i, k + j] == A4[i, j]
    assert A[n - i - 1, k - j - 1] == A1[n - i - 1, k - j - 1]
    assert A[2 * n, 2 * k] == A4[n, k]


def test_block_index_large():
    n, m, k = symbols('n m k', integer=True, positive=True)
    i = symbols('i', integer=True, nonnegative=True)
    A1 = MatrixSymbol('A1', n, n)
    A2 = MatrixSymbol('A2', n, m)
    A3 = MatrixSymbol('A3', n, k)
    A4 = MatrixSymbol('A4', m, n)
    A5 = MatrixSymbol('A5', m, m)
    A6 = MatrixSymbol('A6', m, k)
    A7 = MatrixSymbol('A7', k, n)
    A8 = MatrixSymbol('A8', k, m)
    A9 = MatrixSymbol('A9', k, k)
    A = BlockMatrix([[A1, A2, A3], [A4, A5, A6], [A7, A8, A9]])
    assert A[n + i, n + i] == MatrixElement(A, n + i, n + i)


@XFAIL
def test_block_index_symbolic_fail():
    # To make this work, symbolic matrix dimensions would need to be somehow assumed nonnegative
    # even if the symbols aren't specified as such.  Then 2 * n < n would correctly evaluate to
    # False in BlockMatrix._entry()
    A1 = MatrixSymbol('A1', n, 1)
    A2 = MatrixSymbol('A2', m, 1)
    A = BlockMatrix([[A1], [A2]])
    assert A[2 * n, 0] == A2[n, 0]


def test_slicing():
    A.as_explicit()[0, :]  # does not raise an error


def test_errors():
    raises(IndexError, lambda: Identity(2)[1, 2, 3, 4, 5])
    raises(IndexError, lambda: Identity(2)[[1, 2, 3, 4, 5]])


def test_matrix_expression_to_indices():
    i, j = symbols("i, j")
    i1, i2, i3 = symbols("i_1:4")

    def replace_dummies(expr):
        repl = {i: Symbol(i.name) for i in expr.atoms(Dummy)}
        return expr.xreplace(repl)

    expr = W*X*Z
    assert replace_dummies(expr._entry(i, j)) == \
        Sum(W[i, i1]*X[i1, i2]*Z[i2, j], (i1, 0, l-1), (i2, 0, m-1))
    assert MatrixExpr.from_index_summation(expr._entry(i, j)) == expr

    expr = Z.T*X.T*W.T
    assert replace_dummies(expr._entry(i, j)) == \
        Sum(W[j, i2]*X[i2, i1]*Z[i1, i], (i1, 0, m-1), (i2, 0, l-1))
    assert MatrixExpr.from_index_summation(expr._entry(i, j), i) == expr

    expr = W*X*Z + W*Y*Z
    assert replace_dummies(expr._entry(i, j)) == \
        Sum(W[i, i1]*X[i1, i2]*Z[i2, j], (i1, 0, l-1), (i2, 0, m-1)) +\
        Sum(W[i, i1]*Y[i1, i2]*Z[i2, j], (i1, 0, l-1), (i2, 0, m-1))
    assert MatrixExpr.from_index_summation(expr._entry(i, j)) == expr

    expr = 2*W*X*Z + 3*W*Y*Z
    assert replace_dummies(expr._entry(i, j)) == \
        2*Sum(W[i, i1]*X[i1, i2]*Z[i2, j], (i1, 0, l-1), (i2, 0, m-1)) +\
        3*Sum(W[i, i1]*Y[i1, i2]*Z[i2, j], (i1, 0, l-1), (i2, 0, m-1))
    assert MatrixExpr.from_index_summation(expr._entry(i, j)) == expr

    expr = W*(X + Y)*Z
    assert replace_dummies(expr._entry(i, j)) == \
            Sum(W[i, i1]*(X[i1, i2] + Y[i1, i2])*Z[i2, j], (i1, 0, l-1), (i2, 0, m-1))
    assert MatrixExpr.from_index_summation(expr._entry(i, j)) == expr

    expr = A*B**2*A
    #assert replace_dummies(expr._entry(i, j)) == \
    #        Sum(A[i, i1]*B[i1, i2]*B[i2, i3]*A[i3, j], (i1, 0, 1), (i2, 0, 1), (i3, 0, 1))

    # Check that different dummies are used in sub-multiplications:
    expr = (X1*X2 + X2*X1)*X3
    assert replace_dummies(expr._entry(i, j)) == \
           Sum((Sum(X1[i, i2] * X2[i2, i1], (i2, 0, m - 1)) + Sum(X1[i3, i1] * X2[i, i3], (i3, 0, m - 1))) * X3[
               i1, j], (i1, 0, m - 1))


def test_matrix_expression_from_index_summation():
    from sympy.abc import a,b,c,d
    A = MatrixSymbol("A", k, k)
    B = MatrixSymbol("B", k, k)
    C = MatrixSymbol("C", k, k)
    w1 = MatrixSymbol("w1", k, 1)

    i0, i1, i2, i3, i4 = symbols("i0:5", cls=Dummy)

    expr = Sum(W[a,b]*X[b,c]*Z[c,d], (b, 0, l-1), (c, 0, m-1))
    assert MatrixExpr.from_index_summation(expr, a) == W*X*Z
    expr = Sum(W.T[b,a]*X[b,c]*Z[c,d], (b, 0, l-1), (c, 0, m-1))
    assert MatrixExpr.from_index_summation(expr, a) == W*X*Z
    expr = Sum(A[b, a]*B[b, c]*C[c, d], (b, 0, k-1), (c, 0, k-1))
    assert MatrixSymbol.from_index_summation(expr, a) == A.T*B*C
    expr = Sum(A[b, a]*B[c, b]*C[c, d], (b, 0, k-1), (c, 0, k-1))
    assert MatrixSymbol.from_index_summation(expr, a) == A.T*B.T*C
    expr = Sum(C[c, d]*A[b, a]*B[c, b], (b, 0, k-1), (c, 0, k-1))
    assert MatrixSymbol.from_index_summation(expr, a) == A.T*B.T*C
    expr = Sum(A[a, b] + B[a, b], (a, 0, k-1), (b, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, a) == OneMatrix(1, k)*A*OneMatrix(k, 1) + OneMatrix(1, k)*B*OneMatrix(k, 1)
    expr = Sum(A[a, b]**2, (a, 0, k - 1), (b, 0, k - 1))
    assert MatrixExpr.from_index_summation(expr, a) == Trace(A * A.T)
    expr = Sum(A[a, b]**3, (a, 0, k - 1), (b, 0, k - 1))
    assert MatrixExpr.from_index_summation(expr, a) == Trace(HadamardPower(A.T, 2) * A)
    expr = Sum((A[a, b] + B[a, b])*C[b, c], (b, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, a) == (A+B)*C
    expr = Sum((A[a, b] + B[b, a])*C[b, c], (b, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, a) == (A+B.T)*C
    expr = Sum(A[a, b]*A[b, c]*A[c, d], (b, 0, k-1), (c, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, a) == A**3
    expr = Sum(A[a, b]*A[b, c]*B[c, d], (b, 0, k-1), (c, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, a) == A**2*B

    # Parse the trace of a matrix:

    expr = Sum(A[a, a], (a, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, None) == trace(A)
    expr = Sum(A[a, a]*B[b, c]*C[c, d], (a, 0, k-1), (c, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, b) == trace(A)*B*C

    # Check wrong sum ranges (should raise an exception):

    ## Case 1: 0 to m instead of 0 to m-1
    expr = Sum(W[a,b]*X[b,c]*Z[c,d], (b, 0, l-1), (c, 0, m))
    raises(ValueError, lambda: MatrixExpr.from_index_summation(expr, a))
    ## Case 2: 1 to m-1 instead of 0 to m-1
    expr = Sum(W[a,b]*X[b,c]*Z[c,d], (b, 0, l-1), (c, 1, m-1))
    raises(ValueError, lambda: MatrixExpr.from_index_summation(expr, a))

    # Parse nested sums:
    expr = Sum(A[a, b]*Sum(B[b, c]*C[c, d], (c, 0, k-1)), (b, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, a) == A*B*C

    # Test Kronecker delta:
    expr = Sum(A[a, b]*KroneckerDelta(b, c)*B[c, d], (b, 0, k-1), (c, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, a) == A*B

    expr = Sum(KroneckerDelta(i1, m)*KroneckerDelta(i2, n)*A[i, i1]*A[j, i2], (i1, 0, k-1), (i2, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, m) == ArrayTensorProduct(A.T, A)

    # Test numbered indices:
    expr = Sum(A[i1, i2]*w1[i2, 0], (i2, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, i1) == MatrixElement(A*w1, i1, 0)

    expr = Sum(A[i1, i2]*B[i2, 0], (i2, 0, k-1))
    assert MatrixExpr.from_index_summation(expr, i1) == MatrixElement(A*B, i1, 0)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\io\excel\test_style.py ===
import contextlib
import time

import numpy as np
import pytest

from pandas.compat import is_platform_windows
import pandas.util._test_decorators as td

from pandas import (
    DataFrame,
    read_excel,
)
import pandas._testing as tm

from pandas.io.excel import ExcelWriter
from pandas.io.formats.excel import ExcelFormatter

pytest.importorskip("jinja2")
# jinja2 is currently required for Styler.__init__(). Technically Styler.to_excel
# could compute styles and render to excel without jinja2, since there is no
# 'template' file, but this needs the import error to delayed until render time.

if is_platform_windows():
    pytestmark = pytest.mark.single_cpu


def assert_equal_cell_styles(cell1, cell2):
    # TODO: should find a better way to check equality
    assert cell1.alignment.__dict__ == cell2.alignment.__dict__
    assert cell1.border.__dict__ == cell2.border.__dict__
    assert cell1.fill.__dict__ == cell2.fill.__dict__
    assert cell1.font.__dict__ == cell2.font.__dict__
    assert cell1.number_format == cell2.number_format
    assert cell1.protection.__dict__ == cell2.protection.__dict__


@pytest.mark.parametrize(
    "engine",
    ["xlsxwriter", "openpyxl"],
)
def test_styler_to_excel_unstyled(engine):
    # compare DataFrame.to_excel and Styler.to_excel when no styles applied
    pytest.importorskip(engine)
    df = DataFrame(np.random.default_rng(2).standard_normal((2, 2)))
    with tm.ensure_clean(".xlsx") as path:
        with ExcelWriter(path, engine=engine) as writer:
            df.to_excel(writer, sheet_name="dataframe")
            df.style.to_excel(writer, sheet_name="unstyled")

        openpyxl = pytest.importorskip("openpyxl")  # test loading only with openpyxl
        with contextlib.closing(openpyxl.load_workbook(path)) as wb:
            for col1, col2 in zip(wb["dataframe"].columns, wb["unstyled"].columns):
                assert len(col1) == len(col2)
                for cell1, cell2 in zip(col1, col2):
                    assert cell1.value == cell2.value
                    assert_equal_cell_styles(cell1, cell2)


shared_style_params = [
    (
        "background-color: #111222",
        ["fill", "fgColor", "rgb"],
        {"xlsxwriter": "FF111222", "openpyxl": "00111222"},
    ),
    (
        "color: #111222",
        ["font", "color", "value"],
        {"xlsxwriter": "FF111222", "openpyxl": "00111222"},
    ),
    ("font-family: Arial;", ["font", "name"], "arial"),
    ("font-weight: bold;", ["font", "b"], True),
    ("font-style: italic;", ["font", "i"], True),
    ("text-decoration: underline;", ["font", "u"], "single"),
    ("number-format: $??,???.00;", ["number_format"], "$??,???.00"),
    ("text-align: left;", ["alignment", "horizontal"], "left"),
    (
        "vertical-align: bottom;",
        ["alignment", "vertical"],
        {"xlsxwriter": None, "openpyxl": "bottom"},  # xlsxwriter Fails
    ),
    ("vertical-align: middle;", ["alignment", "vertical"], "center"),
    # Border widths
    ("border-left: 2pt solid red", ["border", "left", "style"], "medium"),
    ("border-left: 1pt dotted red", ["border", "left", "style"], "dotted"),
    ("border-left: 2pt dotted red", ["border", "left", "style"], "mediumDashDotDot"),
    ("border-left: 1pt dashed red", ["border", "left", "style"], "dashed"),
    ("border-left: 2pt dashed red", ["border", "left", "style"], "mediumDashed"),
    ("border-left: 1pt solid red", ["border", "left", "style"], "thin"),
    ("border-left: 3pt solid red", ["border", "left", "style"], "thick"),
    # Border expansion
    (
        "border-left: 2pt solid #111222",
        ["border", "left", "color", "rgb"],
        {"xlsxwriter": "FF111222", "openpyxl": "00111222"},
    ),
    ("border: 1pt solid red", ["border", "top", "style"], "thin"),
    (
        "border: 1pt solid #111222",
        ["border", "top", "color", "rgb"],
        {"xlsxwriter": "FF111222", "openpyxl": "00111222"},
    ),
    ("border: 1pt solid red", ["border", "right", "style"], "thin"),
    (
        "border: 1pt solid #111222",
        ["border", "right", "color", "rgb"],
        {"xlsxwriter": "FF111222", "openpyxl": "00111222"},
    ),
    ("border: 1pt solid red", ["border", "bottom", "style"], "thin"),
    (
        "border: 1pt solid #111222",
        ["border", "bottom", "color", "rgb"],
        {"xlsxwriter": "FF111222", "openpyxl": "00111222"},
    ),
    ("border: 1pt solid red", ["border", "left", "style"], "thin"),
    (
        "border: 1pt solid #111222",
        ["border", "left", "color", "rgb"],
        {"xlsxwriter": "FF111222", "openpyxl": "00111222"},
    ),
    # Border styles
    (
        "border-left-style: hair; border-left-color: black",
        ["border", "left", "style"],
        "hair",
    ),
]


@pytest.mark.parametrize(
    "engine",
    ["xlsxwriter", "openpyxl"],
)
@pytest.mark.parametrize("css, attrs, expected", shared_style_params)
def test_styler_to_excel_basic(engine, css, attrs, expected):
    pytest.importorskip(engine)
    df = DataFrame(np.random.default_rng(2).standard_normal((1, 1)))
    styler = df.style.map(lambda x: css)

    with tm.ensure_clean(".xlsx") as path:
        with ExcelWriter(path, engine=engine) as writer:
            df.to_excel(writer, sheet_name="dataframe")
            styler.to_excel(writer, sheet_name="styled")

        openpyxl = pytest.importorskip("openpyxl")  # test loading only with openpyxl
        with contextlib.closing(openpyxl.load_workbook(path)) as wb:
            # test unstyled data cell does not have expected styles
            # test styled cell has expected styles
            u_cell, s_cell = wb["dataframe"].cell(2, 2), wb["styled"].cell(2, 2)
        for attr in attrs:
            u_cell, s_cell = getattr(u_cell, attr, None), getattr(s_cell, attr)

        if isinstance(expected, dict):
            assert u_cell is None or u_cell != expected[engine]
            assert s_cell == expected[engine]
        else:
            assert u_cell is None or u_cell != expected
            assert s_cell == expected


@pytest.mark.parametrize(
    "engine",
    ["xlsxwriter", "openpyxl"],
)
@pytest.mark.parametrize("css, attrs, expected", shared_style_params)
def test_styler_to_excel_basic_indexes(engine, css, attrs, expected):
    pytest.importorskip(engine)
    df = DataFrame(np.random.default_rng(2).standard_normal((1, 1)))

    styler = df.style
    styler.map_index(lambda x: css, axis=0)
    styler.map_index(lambda x: css, axis=1)

    null_styler = df.style
    null_styler.map(lambda x: "null: css;")
    null_styler.map_index(lambda x: "null: css;", axis=0)
    null_styler.map_index(lambda x: "null: css;", axis=1)

    with tm.ensure_clean(".xlsx") as path:
        with ExcelWriter(path, engine=engine) as writer:
            null_styler.to_excel(writer, sheet_name="null_styled")
            styler.to_excel(writer, sheet_name="styled")

        openpyxl = pytest.importorskip("openpyxl")  # test loading only with openpyxl
        with contextlib.closing(openpyxl.load_workbook(path)) as wb:
            # test null styled index cells does not have expected styles
            # test styled cell has expected styles
            ui_cell, si_cell = wb["null_styled"].cell(2, 1), wb["styled"].cell(2, 1)
            uc_cell, sc_cell = wb["null_styled"].cell(1, 2), wb["styled"].cell(1, 2)
        for attr in attrs:
            ui_cell, si_cell = getattr(ui_cell, attr, None), getattr(si_cell, attr)
            uc_cell, sc_cell = getattr(uc_cell, attr, None), getattr(sc_cell, attr)

        if isinstance(expected, dict):
            assert ui_cell is None or ui_cell != expected[engine]
            assert si_cell == expected[engine]
            assert uc_cell is None or uc_cell != expected[engine]
            assert sc_cell == expected[engine]
        else:
            assert ui_cell is None or ui_cell != expected
            assert si_cell == expected
            assert uc_cell is None or uc_cell != expected
            assert sc_cell == expected


# From https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.borders.html
# Note: Leaving behavior of "width"-type styles undefined; user should use border-width
# instead
excel_border_styles = [
    # "thin",
    "dashed",
    "mediumDashDot",
    "dashDotDot",
    "hair",
    "dotted",
    "mediumDashDotDot",
    # "medium",
    "double",
    "dashDot",
    "slantDashDot",
    # "thick",
    "mediumDashed",
]


@pytest.mark.parametrize(
    "engine",
    ["xlsxwriter", "openpyxl"],
)
@pytest.mark.parametrize("border_style", excel_border_styles)
def test_styler_to_excel_border_style(engine, border_style):
    css = f"border-left: {border_style} black thin"
    attrs = ["border", "left", "style"]
    expected = border_style

    pytest.importorskip(engine)
    df = DataFrame(np.random.default_rng(2).standard_normal((1, 1)))
    styler = df.style.map(lambda x: css)

    with tm.ensure_clean(".xlsx") as path:
        with ExcelWriter(path, engine=engine) as writer:
            df.to_excel(writer, sheet_name="dataframe")
            styler.to_excel(writer, sheet_name="styled")

        openpyxl = pytest.importorskip("openpyxl")  # test loading only with openpyxl
        with contextlib.closing(openpyxl.load_workbook(path)) as wb:
            # test unstyled data cell does not have expected styles
            # test styled cell has expected styles
            u_cell, s_cell = wb["dataframe"].cell(2, 2), wb["styled"].cell(2, 2)
        for attr in attrs:
            u_cell, s_cell = getattr(u_cell, attr, None), getattr(s_cell, attr)

        if isinstance(expected, dict):
            assert u_cell is None or u_cell != expected[engine]
            assert s_cell == expected[engine]
        else:
            assert u_cell is None or u_cell != expected
            assert s_cell == expected


def test_styler_custom_converter():
    openpyxl = pytest.importorskip("openpyxl")

    def custom_converter(css):
        return {"font": {"color": {"rgb": "111222"}}}

    df = DataFrame(np.random.default_rng(2).standard_normal((1, 1)))
    styler = df.style.map(lambda x: "color: #888999")
    with tm.ensure_clean(".xlsx") as path:
        with ExcelWriter(path, engine="openpyxl") as writer:
            ExcelFormatter(styler, style_converter=custom_converter).write(
                writer, sheet_name="custom"
            )

        with contextlib.closing(openpyxl.load_workbook(path)) as wb:
            assert wb["custom"].cell(2, 2).font.color.value == "00111222"


@pytest.mark.single_cpu
@td.skip_if_not_us_locale
def test_styler_to_s3(s3_public_bucket, s3so):
    # GH#46381

    mock_bucket_name, target_file = s3_public_bucket.name, "test.xlsx"
    df = DataFrame({"x": [1, 2, 3], "y": [2, 4, 6]})
    styler = df.style.set_sticky(axis="index")
    styler.to_excel(f"s3://{mock_bucket_name}/{target_file}", storage_options=s3so)
    timeout = 5
    while True:
        if target_file in (obj.key for obj in s3_public_bucket.objects.all()):
            break
        time.sleep(0.1)
        timeout -= 0.1
        assert timeout > 0, "Timed out waiting for file to appear on moto"
        result = read_excel(
            f"s3://{mock_bucket_name}/{target_file}", index_col=0, storage_options=s3so
        )
        tm.assert_frame_equal(result, df)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\numberfields\tests\test_primes.py ===
from math import prod

from sympy import QQ, ZZ
from sympy.abc import x, theta
from sympy.ntheory import factorint
from sympy.ntheory.residue_ntheory import n_order
from sympy.polys import Poly, cyclotomic_poly
from sympy.polys.matrices import DomainMatrix
from sympy.polys.numberfields.basis import round_two
from sympy.polys.numberfields.exceptions import StructureError
from sympy.polys.numberfields.modules import PowerBasis, to_col
from sympy.polys.numberfields.primes import (
    prime_decomp, _two_elt_rep,
    _check_formal_conditions_for_maximal_order,
)
from sympy.testing.pytest import raises


def test_check_formal_conditions_for_maximal_order():
    T = Poly(cyclotomic_poly(5, x))
    A = PowerBasis(T)
    B = A.submodule_from_matrix(2 * DomainMatrix.eye(4, ZZ))
    C = B.submodule_from_matrix(3 * DomainMatrix.eye(4, ZZ))
    D = A.submodule_from_matrix(DomainMatrix.eye(4, ZZ)[:, :-1])
    # Is a direct submodule of a power basis, but lacks 1 as first generator:
    raises(StructureError, lambda: _check_formal_conditions_for_maximal_order(B))
    # Is not a direct submodule of a power basis:
    raises(StructureError, lambda: _check_formal_conditions_for_maximal_order(C))
    # Is direct submod of pow basis, and starts with 1, but not sq/max rank/HNF:
    raises(StructureError, lambda: _check_formal_conditions_for_maximal_order(D))


def test_two_elt_rep():
    ell = 7
    T = Poly(cyclotomic_poly(ell))
    ZK, dK = round_two(T)
    for p in [29, 13, 11, 5]:
        P = prime_decomp(p, T)
        for Pi in P:
            # We have Pi in two-element representation, and, because we are
            # looking at a cyclotomic field, this was computed by the "easy"
            # method that just factors T mod p. We will now convert this to
            # a set of Z-generators, then convert that back into a two-element
            # rep. The latter need not be identical to the two-elt rep we
            # already have, but it must have the same HNF.
            H = p*ZK + Pi.alpha*ZK
            gens = H.basis_element_pullbacks()
            # Note: we could supply f = Pi.f, but prefer to test behavior without it.
            b = _two_elt_rep(gens, ZK, p)
            if b != Pi.alpha:
                H2 = p*ZK + b*ZK
                assert H2 == H


def test_valuation_at_prime_ideal():
    p = 7
    T = Poly(cyclotomic_poly(p))
    ZK, dK = round_two(T)
    P = prime_decomp(p, T, dK=dK, ZK=ZK)
    assert len(P) == 1
    P0 = P[0]
    v = P0.valuation(p*ZK)
    assert v == P0.e
    # Test easy 0 case:
    assert P0.valuation(5*ZK) == 0


def test_decomp_1():
    # All prime decompositions in cyclotomic fields are in the "easy case,"
    # since the index is unity.
    # Here we check the ramified prime.
    T = Poly(cyclotomic_poly(7))
    raises(ValueError, lambda: prime_decomp(7))
    P = prime_decomp(7, T)
    assert len(P) == 1
    P0 = P[0]
    assert P0.e == 6
    assert P0.f == 1
    # Test powers:
    assert P0**0 == P0.ZK
    assert P0**1 == P0
    assert P0**6 == 7 * P0.ZK


def test_decomp_2():
    # More easy cyclotomic cases, but here we check unramified primes.
    ell = 7
    T = Poly(cyclotomic_poly(ell))
    for p in [29, 13, 11, 5]:
        f_exp = n_order(p, ell)
        g_exp = (ell - 1) // f_exp
        P = prime_decomp(p, T)
        assert len(P) == g_exp
        for Pi in P:
            assert Pi.e == 1
            assert Pi.f == f_exp


def test_decomp_3():
    T = Poly(x ** 2 - 35)
    rad = {}
    ZK, dK = round_two(T, radicals=rad)
    # 35 is 3 mod 4, so field disc is 4*5*7, and theory says each of the
    # rational primes 2, 5, 7 should be the square of a prime ideal.
    for p in [2, 5, 7]:
        P = prime_decomp(p, T, dK=dK, ZK=ZK, radical=rad.get(p))
        assert len(P) == 1
        assert P[0].e == 2
        assert P[0]**2 == p*ZK


def test_decomp_4():
    T = Poly(x ** 2 - 21)
    rad = {}
    ZK, dK = round_two(T, radicals=rad)
    # 21 is 1 mod 4, so field disc is 3*7, and theory says the
    # rational primes 3, 7 should be the square of a prime ideal.
    for p in [3, 7]:
        P = prime_decomp(p, T, dK=dK, ZK=ZK, radical=rad.get(p))
        assert len(P) == 1
        assert P[0].e == 2
        assert P[0]**2 == p*ZK


def test_decomp_5():
    # Here is our first test of the "hard case" of prime decomposition.
    # We work in a quadratic extension Q(sqrt(d)) where d is 1 mod 4, and
    # we consider the factorization of the rational prime 2, which divides
    # the index.
    # Theory says the form of p's factorization depends on the residue of
    # d mod 8, so we consider both cases, d = 1 mod 8 and d = 5 mod 8.
    for d in [-7, -3]:
        T = Poly(x ** 2 - d)
        rad = {}
        ZK, dK = round_two(T, radicals=rad)
        p = 2
        P = prime_decomp(p, T, dK=dK, ZK=ZK, radical=rad.get(p))
        if d % 8 == 1:
            assert len(P) == 2
            assert all(P[i].e == 1 and P[i].f == 1 for i in range(2))
            assert prod(Pi**Pi.e for Pi in P) == p * ZK
        else:
            assert d % 8 == 5
            assert len(P) == 1
            assert P[0].e == 1
            assert P[0].f == 2
            assert P[0].as_submodule() == p * ZK


def test_decomp_6():
    # Another case where 2 divides the index. This is Dedekind's example of
    # an essential discriminant divisor. (See Cohen, Exercise 6.10.)
    T = Poly(x ** 3 + x ** 2 - 2 * x + 8)
    rad = {}
    ZK, dK = round_two(T, radicals=rad)
    p = 2
    P = prime_decomp(p, T, dK=dK, ZK=ZK, radical=rad.get(p))
    assert len(P) == 3
    assert all(Pi.e == Pi.f == 1 for Pi in P)
    assert prod(Pi**Pi.e for Pi in P) == p*ZK


def test_decomp_7():
    # Try working through an AlgebraicField
    T = Poly(x ** 3 + x ** 2 - 2 * x + 8)
    K = QQ.alg_field_from_poly(T)
    p = 2
    P = K.primes_above(p)
    ZK = K.maximal_order()
    assert len(P) == 3
    assert all(Pi.e == Pi.f == 1 for Pi in P)
    assert prod(Pi**Pi.e for Pi in P) == p*ZK


def test_decomp_8():
    # This time we consider various cubics, and try factoring all primes
    # dividing the index.
    cases = (
        x ** 3 + 3 * x ** 2 - 4 * x + 4,
        x ** 3 + 3 * x ** 2 + 3 * x - 3,
        x ** 3 + 5 * x ** 2 - x + 3,
        x ** 3 + 5 * x ** 2 - 5 * x - 5,
        x ** 3 + 3 * x ** 2 + 5,
        x ** 3 + 6 * x ** 2 + 3 * x - 1,
        x ** 3 + 6 * x ** 2 + 4,
        x ** 3 + 7 * x ** 2 + 7 * x - 7,
        x ** 3 + 7 * x ** 2 - x + 5,
        x ** 3 + 7 * x ** 2 - 5 * x + 5,
        x ** 3 + 4 * x ** 2 - 3 * x + 7,
        x ** 3 + 8 * x ** 2 + 5 * x - 1,
        x ** 3 + 8 * x ** 2 - 2 * x + 6,
        x ** 3 + 6 * x ** 2 - 3 * x + 8,
        x ** 3 + 9 * x ** 2 + 6 * x - 8,
        x ** 3 + 15 * x ** 2 - 9 * x + 13,
    )
    def display(T, p, radical, P, I, J):
        """Useful for inspection, when running test manually."""
        print('=' * 20)
        print(T, p, radical)
        for Pi in P:
            print(f'  ({Pi!r})')
        print("I: ", I)
        print("J: ", J)
        print(f'Equal: {I == J}')
    inspect = False
    for g in cases:
        T = Poly(g)
        rad = {}
        ZK, dK = round_two(T, radicals=rad)
        dT = T.discriminant()
        f_squared = dT // dK
        F = factorint(f_squared)
        for p in F:
            radical = rad.get(p)
            P = prime_decomp(p, T, dK=dK, ZK=ZK, radical=radical)
            I = prod(Pi**Pi.e for Pi in P)
            J = p * ZK
            if inspect:
                display(T, p, radical, P, I, J)
            assert I == J


def test_PrimeIdeal_eq():
    # `==` should fail on objects of different types, so even a completely
    # inert PrimeIdeal should test unequal to the rational prime it divides.
    T = Poly(cyclotomic_poly(7))
    P0 = prime_decomp(5, T)[0]
    assert P0.f == 6
    assert P0.as_submodule() == 5 * P0.ZK
    assert P0 != 5


def test_PrimeIdeal_add():
    T = Poly(cyclotomic_poly(7))
    P0 = prime_decomp(7, T)[0]
    # Adding ideals computes their GCD, so adding the ramified prime dividing
    # 7 to 7 itself should reproduce this prime (as a submodule).
    assert P0 + 7 * P0.ZK == P0.as_submodule()


def test_str():
    # Without alias:
    k = QQ.alg_field_from_poly(Poly(x**2 + 7))
    frp = k.primes_above(2)[0]
    assert str(frp) == '(2, 3*_x/2 + 1/2)'

    frp = k.primes_above(3)[0]
    assert str(frp) == '(3)'

    # With alias:
    k = QQ.alg_field_from_poly(Poly(x ** 2 + 7), alias='alpha')
    frp = k.primes_above(2)[0]
    assert str(frp) == '(2, 3*alpha/2 + 1/2)'

    frp = k.primes_above(3)[0]
    assert str(frp) == '(3)'


def test_repr():
    T = Poly(x**2 + 7)
    ZK, dK = round_two(T)
    P = prime_decomp(2, T, dK=dK, ZK=ZK)
    assert repr(P[0]) == '[ (2, (3*x + 1)/2) e=1, f=1 ]'
    assert P[0].repr(field_gen=theta) == '[ (2, (3*theta + 1)/2) e=1, f=1 ]'
    assert P[0].repr(field_gen=theta, just_gens=True) == '(2, (3*theta + 1)/2)'


def test_PrimeIdeal_reduce():
    k = QQ.alg_field_from_poly(Poly(x ** 3 + x ** 2 - 2 * x + 8))
    Zk = k.maximal_order()
    P = k.primes_above(2)
    frp = P[2]

    # reduce_element
    a = Zk.parent(to_col([23, 20, 11]), denom=6)
    a_bar_expected = Zk.parent(to_col([11, 5, 2]), denom=6)
    a_bar = frp.reduce_element(a)
    assert a_bar == a_bar_expected

    # reduce_ANP
    a = k([QQ(11, 6), QQ(20, 6), QQ(23, 6)])
    a_bar_expected = k([QQ(2, 6), QQ(5, 6), QQ(11, 6)])
    a_bar = frp.reduce_ANP(a)
    assert a_bar == a_bar_expected

    # reduce_alg_num
    a = k.to_alg_num(a)
    a_bar_expected = k.to_alg_num(a_bar_expected)
    a_bar = frp.reduce_alg_num(a)
    assert a_bar == a_bar_expected


def test_issue_23402():
    k = QQ.alg_field_from_poly(Poly(x ** 3 + x ** 2 - 2 * x + 8))
    P = k.primes_above(3)
    assert P[0].alpha.equiv(0)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\solvers\tests\test_recurr.py ===
from sympy.core.function import (Function, Lambda, expand)
from sympy.core.numbers import (I, Rational)
from sympy.core.relational import Eq
from sympy.core.singleton import S
from sympy.core.symbol import (Symbol, symbols)
from sympy.functions.combinatorial.factorials import (rf, binomial, factorial)
from sympy.functions.elementary.complexes import Abs
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.trigonometric import (cos, sin)
from sympy.polys.polytools import factor
from sympy.solvers.recurr import rsolve, rsolve_hyper, rsolve_poly, rsolve_ratio
from sympy.testing.pytest import raises, slow, XFAIL
from sympy.abc import a, b

y = Function('y')
n, k = symbols('n,k', integer=True)
C0, C1, C2 = symbols('C0,C1,C2')


def test_rsolve_poly():
    assert rsolve_poly([-1, -1, 1], 0, n) == 0
    assert rsolve_poly([-1, -1, 1], 1, n) == -1

    assert rsolve_poly([-1, n + 1], n, n) == 1
    assert rsolve_poly([-1, 1], n, n) == C0 + (n**2 - n)/2
    assert rsolve_poly([-n - 1, n], 1, n) == C0*n - 1
    assert rsolve_poly([-4*n - 2, 1], 4*n + 1, n) == -1

    assert rsolve_poly([-1, 1], n**5 + n**3, n) == \
        C0 - n**3 / 2 - n**5 / 2 + n**2 / 6 + n**6 / 6 + 2*n**4 / 3


def test_rsolve_ratio():
    solution = rsolve_ratio([-2*n**3 + n**2 + 2*n - 1, 2*n**3 + n**2 - 6*n,
        -2*n**3 - 11*n**2 - 18*n - 9, 2*n**3 + 13*n**2 + 22*n + 8], 0, n)
    assert solution == C0*(2*n - 3)/(n**2 - 1)/2


def test_rsolve_hyper():
    assert rsolve_hyper([-1, -1, 1], 0, n) in [
        C0*(S.Half - S.Half*sqrt(5))**n + C1*(S.Half + S.Half*sqrt(5))**n,
        C1*(S.Half - S.Half*sqrt(5))**n + C0*(S.Half + S.Half*sqrt(5))**n,
    ]

    assert rsolve_hyper([n**2 - 2, -2*n - 1, 1], 0, n) in [
        C0*rf(sqrt(2), n) + C1*rf(-sqrt(2), n),
        C1*rf(sqrt(2), n) + C0*rf(-sqrt(2), n),
    ]

    assert rsolve_hyper([n**2 - k, -2*n - 1, 1], 0, n) in [
        C0*rf(sqrt(k), n) + C1*rf(-sqrt(k), n),
        C1*rf(sqrt(k), n) + C0*rf(-sqrt(k), n),
    ]

    assert rsolve_hyper(
        [2*n*(n + 1), -n**2 - 3*n + 2, n - 1], 0, n) == C1*factorial(n) + C0*2**n

    assert rsolve_hyper(
        [n + 2, -(2*n + 3)*(17*n**2 + 51*n + 39), n + 1], 0, n) == 0

    assert rsolve_hyper([-n - 1, -1, 1], 0, n) == 0

    assert rsolve_hyper([-1, 1], n, n).expand() == C0 + n**2/2 - n/2

    assert rsolve_hyper([-1, 1], 1 + n, n).expand() == C0 + n**2/2 + n/2

    assert rsolve_hyper([-1, 1], 3*(n + n**2), n).expand() == C0 + n**3 - n

    assert rsolve_hyper([-a, 1],0,n).expand() == C0*a**n

    assert rsolve_hyper([-a, 0, 1], 0, n).expand() == (-1)**n*C1*a**(n/2) + C0*a**(n/2)

    assert rsolve_hyper([1, 1, 1], 0, n).expand() == \
        C0*(Rational(-1, 2) - sqrt(3)*I/2)**n + C1*(Rational(-1, 2) + sqrt(3)*I/2)**n

    assert rsolve_hyper([1, -2*n/a - 2/a, 1], 0, n) == 0


@XFAIL
def test_rsolve_ratio_missed():
    # this arises during computation
    # assert rsolve_hyper([-1, 1], 3*(n + n**2), n).expand() == C0 + n**3 - n
    assert rsolve_ratio([-n, n + 2], n, n) is not None


def recurrence_term(c, f):
    """Compute RHS of recurrence in f(n) with coefficients in c."""
    return sum(c[i]*f.subs(n, n + i) for i in range(len(c)))


def test_rsolve_bulk():
    """Some bulk-generated tests."""
    funcs = [ n, n + 1, n**2, n**3, n**4, n + n**2, 27*n + 52*n**2 - 3*
        n**3 + 12*n**4 - 52*n**5 ]
    coeffs = [ [-2, 1], [-2, -1, 1], [-1, 1, 1, -1, 1], [-n, 1], [n**2 -
        n + 12, 1] ]
    for p in funcs:
        # compute difference
        for c in coeffs:
            q = recurrence_term(c, p)
            if p.is_polynomial(n):
                assert rsolve_poly(c, q, n) == p
            # See issue 3956:
            if p.is_hypergeometric(n) and len(c) <= 3:
                assert rsolve_hyper(c, q, n).subs(zip(symbols('C:3'), [0, 0, 0])).expand() == p


def test_rsolve_0_sol_homogeneous():
    # fixed by cherry-pick from
    # https://github.com/diofant/diofant/commit/e1d2e52125199eb3df59f12e8944f8a5f24b00a5
    assert rsolve_hyper([n**2 - n + 12, 1], n*(n**2 - n + 12) + n + 1, n) == n


def test_rsolve():
    f = y(n + 2) - y(n + 1) - y(n)
    h = sqrt(5)*(S.Half + S.Half*sqrt(5))**n \
        - sqrt(5)*(S.Half - S.Half*sqrt(5))**n

    assert rsolve(f, y(n)) in [
        C0*(S.Half - S.Half*sqrt(5))**n + C1*(S.Half + S.Half*sqrt(5))**n,
        C1*(S.Half - S.Half*sqrt(5))**n + C0*(S.Half + S.Half*sqrt(5))**n,
    ]

    assert rsolve(f, y(n), [0, 5]) == h
    assert rsolve(f, y(n), {0: 0, 1: 5}) == h
    assert rsolve(f, y(n), {y(0): 0, y(1): 5}) == h
    assert rsolve(y(n) - y(n - 1) - y(n - 2), y(n), [0, 5]) == h
    assert rsolve(Eq(y(n), y(n - 1) + y(n - 2)), y(n), [0, 5]) == h

    assert f.subs(y, Lambda(k, rsolve(f, y(n)).subs(n, k))).simplify() == 0

    f = (n - 1)*y(n + 2) - (n**2 + 3*n - 2)*y(n + 1) + 2*n*(n + 1)*y(n)
    g = C1*factorial(n) + C0*2**n
    h = -3*factorial(n) + 3*2**n

    assert rsolve(f, y(n)) == g
    assert rsolve(f, y(n), []) == g
    assert rsolve(f, y(n), {}) == g

    assert rsolve(f, y(n), [0, 3]) == h
    assert rsolve(f, y(n), {0: 0, 1: 3}) == h
    assert rsolve(f, y(n), {y(0): 0, y(1): 3}) == h

    assert f.subs(y, Lambda(k, rsolve(f, y(n)).subs(n, k))).simplify() == 0

    f = y(n) - y(n - 1) - 2

    assert rsolve(f, y(n), {y(0): 0}) == 2*n
    assert rsolve(f, y(n), {y(0): 1}) == 2*n + 1
    assert rsolve(f, y(n), {y(0): 0, y(1): 1}) is None

    assert f.subs(y, Lambda(k, rsolve(f, y(n)).subs(n, k))).simplify() == 0

    f = 3*y(n - 1) - y(n) - 1

    assert rsolve(f, y(n), {y(0): 0}) == -3**n/2 + S.Half
    assert rsolve(f, y(n), {y(0): 1}) == 3**n/2 + S.Half
    assert rsolve(f, y(n), {y(0): 2}) == 3*3**n/2 + S.Half

    assert f.subs(y, Lambda(k, rsolve(f, y(n)).subs(n, k))).simplify() == 0

    f = y(n) - 1/n*y(n - 1)
    assert rsolve(f, y(n)) == C0/factorial(n)
    assert f.subs(y, Lambda(k, rsolve(f, y(n)).subs(n, k))).simplify() == 0

    f = y(n) - 1/n*y(n - 1) - 1
    assert rsolve(f, y(n)) is None

    f = 2*y(n - 1) + (1 - n)*y(n)/n

    assert rsolve(f, y(n), {y(1): 1}) == 2**(n - 1)*n
    assert rsolve(f, y(n), {y(1): 2}) == 2**(n - 1)*n*2
    assert rsolve(f, y(n), {y(1): 3}) == 2**(n - 1)*n*3

    assert f.subs(y, Lambda(k, rsolve(f, y(n)).subs(n, k))).simplify() == 0

    f = (n - 1)*(n - 2)*y(n + 2) - (n + 1)*(n + 2)*y(n)

    assert rsolve(f, y(n), {y(3): 6, y(4): 24}) == n*(n - 1)*(n - 2)
    assert rsolve(
        f, y(n), {y(3): 6, y(4): -24}) == -n*(n - 1)*(n - 2)*(-1)**(n)

    assert f.subs(y, Lambda(k, rsolve(f, y(n)).subs(n, k))).simplify() == 0

    assert rsolve(Eq(y(n + 1), a*y(n)), y(n), {y(1): a}).simplify() == a**n

    assert rsolve(y(n) - a*y(n-2),y(n), \
            {y(1): sqrt(a)*(a + b), y(2): a*(a - b)}).simplify() == \
            a**(n/2 + 1) - b*(-sqrt(a))**n

    f = (-16*n**2 + 32*n - 12)*y(n - 1) + (4*n**2 - 12*n + 9)*y(n)

    yn = rsolve(f, y(n), {y(1): binomial(2*n + 1, 3)})
    sol = 2**(2*n)*n*(2*n - 1)**2*(2*n + 1)/12
    assert factor(expand(yn, func=True)) == sol

    sol = rsolve(y(n) + a*(y(n + 1) + y(n - 1))/2, y(n))
    assert str(sol) == 'C0*((-sqrt(1 - a**2) - 1)/a)**n + C1*((sqrt(1 - a**2) - 1)/a)**n'

    assert rsolve((k + 1)*y(k), y(k)) is None
    assert (rsolve((k + 1)*y(k) + (k + 3)*y(k + 1) + (k + 5)*y(k + 2), y(k))
            is None)

    assert rsolve(y(n) + y(n + 1) + 2**n + 3**n, y(n)) == (-1)**n*C0 - 2**n/3 - 3**n/4


def test_rsolve_raises():
    x = Function('x')
    raises(ValueError, lambda: rsolve(y(n) - y(k + 1), y(n)))
    raises(ValueError, lambda: rsolve(y(n) - y(n + 1), x(n)))
    raises(ValueError, lambda: rsolve(y(n) - x(n + 1), y(n)))
    raises(ValueError, lambda: rsolve(y(n) - sqrt(n)*y(n + 1), y(n)))
    raises(ValueError, lambda: rsolve(y(n) - y(n + 1), y(n), {x(0): 0}))
    raises(ValueError, lambda: rsolve(y(n) + y(n + 1) + 2**n + cos(n), y(n)))


def test_issue_6844():
    f = y(n + 2) - y(n + 1) + y(n)/4
    assert rsolve(f, y(n)) == 2**(-n + 1)*C1*n + 2**(-n)*C0
    assert rsolve(f, y(n), {y(0): 0, y(1): 1}) == 2**(1 - n)*n


def test_issue_18751():
    r = Symbol('r', positive=True)
    theta = Symbol('theta', real=True)
    f = y(n) - 2 * r * cos(theta) * y(n - 1) + r**2 * y(n - 2)
    assert rsolve(f, y(n)) == \
        C0*(r*(cos(theta) - I*Abs(sin(theta))))**n + C1*(r*(cos(theta) + I*Abs(sin(theta))))**n


def test_constant_naming():
    #issue 8697
    assert rsolve(y(n+3) - y(n+2) - y(n+1) + y(n), y(n)) == (-1)**n*C1 + C0 + C2*n
    assert rsolve(y(n+3)+3*y(n+2)+3*y(n+1)+y(n), y(n)).expand() == (-1)**n*C0 - (-1)**n*C1*n - (-1)**n*C2*n**2
    assert rsolve(y(n) - 2*y(n - 3) + 5*y(n - 2) - 4*y(n - 1),y(n),[1,3,8]) == 3*2**n - n - 2

    #issue 19630
    assert rsolve(y(n+3) - 3*y(n+1) + 2*y(n), y(n), {y(1):0, y(2):8, y(3):-2}) == (-2)**n + 2*n


@slow
def test_issue_15751():
    f = y(n) + 21*y(n + 1) - 273*y(n + 2) - 1092*y(n + 3) + 1820*y(n + 4) + 1092*y(n + 5) - 273*y(n + 6) - 21*y(n + 7) + y(n + 8)
    assert rsolve(f, y(n)) is not None


def test_issue_17990():
    f = -10*y(n) + 4*y(n + 1) + 6*y(n + 2) + 46*y(n + 3)
    sol = rsolve(f, y(n))
    expected = C0*((86*18**(S(1)/3)/69 + (-12 + (-1 + sqrt(3)*I)*(290412 +
        3036*sqrt(9165))**(S(1)/3))*(1 - sqrt(3)*I)*(24201 + 253*sqrt(9165))**
        (S(1)/3)/276)/((1 - sqrt(3)*I)*(24201 + 253*sqrt(9165))**(S(1)/3))
        )**n + C1*((86*18**(S(1)/3)/69 + (-12 + (-1 - sqrt(3)*I)*(290412 + 3036
        *sqrt(9165))**(S(1)/3))*(1 + sqrt(3)*I)*(24201 + 253*sqrt(9165))**
        (S(1)/3)/276)/((1 + sqrt(3)*I)*(24201 + 253*sqrt(9165))**(S(1)/3))
        )**n + C2*(-43*18**(S(1)/3)/(69*(24201 + 253*sqrt(9165))**(S(1)/3)) -
        S(1)/23 + (290412 + 3036*sqrt(9165))**(S(1)/3)/138)**n
    assert sol == expected
    e = sol.subs({C0: 1, C1: 1, C2: 1, n: 1}).evalf()
    assert abs(e + 0.130434782608696) < 1e-13


def test_issue_8697():
    a = Function('a')
    eq = a(n + 3) - a(n + 2) - a(n + 1) + a(n)
    assert rsolve(eq, a(n)) == (-1)**n*C1 + C0 + C2*n
    eq2 = a(n + 3) + 3*a(n + 2) + 3*a(n + 1) + a(n)
    assert (rsolve(eq2, a(n)) ==
            (-1)**n*C0 + (-1)**(n + 1)*C1*n + (-1)**(n + 1)*C2*n**2)

    assert rsolve(a(n) - 2*a(n - 3) + 5*a(n - 2) - 4*a(n - 1),
                  a(n), {a(0): 1, a(1): 3, a(2): 8}) == 3*2**n - n - 2

    # From issue thread (but fixed by https://github.com/diofant/diofant/commit/da9789c6cd7d0c2ceeea19fbf59645987125b289):
    assert rsolve(a(n) - 2*a(n - 1) - n, a(n), {a(0): 1}) == 3*2**n - n - 2


def test_diofantissue_294():
    f = y(n) - y(n - 1) - 2*y(n - 2) - 2*n
    assert rsolve(f, y(n)) == (-1)**n*C0 + 2**n*C1 - n - Rational(5, 2)
    # issue sympy/sympy#11261
    assert rsolve(f, y(n), {y(0): -1, y(1): 1}) == (-(-1)**n/2 + 2*2**n -
                                                    n - Rational(5, 2))
    # issue sympy/sympy#7055
    assert rsolve(-2*y(n) + y(n + 1) + n - 1, y(n)) == 2**n*C0 + n


def test_issue_15553():
    f = Function("f")
    assert rsolve(Eq(f(n), 2*f(n - 1) + n), f(n)) == 2**n*C0 - n - 2
    assert rsolve(Eq(f(n + 1), 2*f(n) + n**2 + 1), f(n)) == 2**n*C0 - n**2 - 2*n - 4
    assert rsolve(Eq(f(n + 1), 2*f(n) + n**2 + 1), f(n), {f(1): 0}) == 7*2**n/2 - n**2 - 2*n - 4
    assert rsolve(Eq(f(n), 2*f(n - 1) + 3*n**2), f(n)) == 2**n*C0 - 3*n**2 - 12*n - 18
    assert rsolve(Eq(f(n), 2*f(n - 1) + n**2), f(n)) == 2**n*C0 - n**2 - 4*n - 6
    assert rsolve(Eq(f(n), 2*f(n - 1) + n), f(n), {f(0): 1}) == 3*2**n - n - 2

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\colorama\tests\ansitowin32_test.py ===
# Copyright Jonathan Hartley 2013. BSD 3-Clause license, see LICENSE file.
from io import StringIO, TextIOWrapper
from unittest import TestCase, main
try:
    from contextlib import ExitStack
except ImportError:
    # python 2
    from contextlib2 import ExitStack

try:
    from unittest.mock import MagicMock, Mock, patch
except ImportError:
    from mock import MagicMock, Mock, patch

from ..ansitowin32 import AnsiToWin32, StreamWrapper
from ..win32 import ENABLE_VIRTUAL_TERMINAL_PROCESSING
from .utils import osname


class StreamWrapperTest(TestCase):

    def testIsAProxy(self):
        mockStream = Mock()
        wrapper = StreamWrapper(mockStream, None)
        self.assertTrue( wrapper.random_attr is mockStream.random_attr )

    def testDelegatesWrite(self):
        mockStream = Mock()
        mockConverter = Mock()
        wrapper = StreamWrapper(mockStream, mockConverter)
        wrapper.write('hello')
        self.assertTrue(mockConverter.write.call_args, (('hello',), {}))

    def testDelegatesContext(self):
        mockConverter = Mock()
        s = StringIO()
        with StreamWrapper(s, mockConverter) as fp:
            fp.write(u'hello')
        self.assertTrue(s.closed)

    def testProxyNoContextManager(self):
        mockStream = MagicMock()
        mockStream.__enter__.side_effect = AttributeError()
        mockConverter = Mock()
        with self.assertRaises(AttributeError) as excinfo:
            with StreamWrapper(mockStream, mockConverter) as wrapper:
                wrapper.write('hello')

    def test_closed_shouldnt_raise_on_closed_stream(self):
        stream = StringIO()
        stream.close()
        wrapper = StreamWrapper(stream, None)
        self.assertEqual(wrapper.closed, True)

    def test_closed_shouldnt_raise_on_detached_stream(self):
        stream = TextIOWrapper(StringIO())
        stream.detach()
        wrapper = StreamWrapper(stream, None)
        self.assertEqual(wrapper.closed, True)

class AnsiToWin32Test(TestCase):

    def testInit(self):
        mockStdout = Mock()
        auto = Mock()
        stream = AnsiToWin32(mockStdout, autoreset=auto)
        self.assertEqual(stream.wrapped, mockStdout)
        self.assertEqual(stream.autoreset, auto)

    @patch('colorama.ansitowin32.winterm', None)
    @patch('colorama.ansitowin32.winapi_test', lambda *_: True)
    def testStripIsTrueOnWindows(self):
        with osname('nt'):
            mockStdout = Mock()
            stream = AnsiToWin32(mockStdout)
            self.assertTrue(stream.strip)

    def testStripIsFalseOffWindows(self):
        with osname('posix'):
            mockStdout = Mock(closed=False)
            stream = AnsiToWin32(mockStdout)
            self.assertFalse(stream.strip)

    def testWriteStripsAnsi(self):
        mockStdout = Mock()
        stream = AnsiToWin32(mockStdout)
        stream.wrapped = Mock()
        stream.write_and_convert = Mock()
        stream.strip = True

        stream.write('abc')

        self.assertFalse(stream.wrapped.write.called)
        self.assertEqual(stream.write_and_convert.call_args, (('abc',), {}))

    def testWriteDoesNotStripAnsi(self):
        mockStdout = Mock()
        stream = AnsiToWin32(mockStdout)
        stream.wrapped = Mock()
        stream.write_and_convert = Mock()
        stream.strip = False
        stream.convert = False

        stream.write('abc')

        self.assertFalse(stream.write_and_convert.called)
        self.assertEqual(stream.wrapped.write.call_args, (('abc',), {}))

    def assert_autoresets(self, convert, autoreset=True):
        stream = AnsiToWin32(Mock())
        stream.convert = convert
        stream.reset_all = Mock()
        stream.autoreset = autoreset
        stream.winterm = Mock()

        stream.write('abc')

        self.assertEqual(stream.reset_all.called, autoreset)

    def testWriteAutoresets(self):
        self.assert_autoresets(convert=True)
        self.assert_autoresets(convert=False)
        self.assert_autoresets(convert=True, autoreset=False)
        self.assert_autoresets(convert=False, autoreset=False)

    def testWriteAndConvertWritesPlainText(self):
        stream = AnsiToWin32(Mock())
        stream.write_and_convert( 'abc' )
        self.assertEqual( stream.wrapped.write.call_args, (('abc',), {}) )

    def testWriteAndConvertStripsAllValidAnsi(self):
        stream = AnsiToWin32(Mock())
        stream.call_win32 = Mock()
        data = [
            'abc\033[mdef',
            'abc\033[0mdef',
            'abc\033[2mdef',
            'abc\033[02mdef',
            'abc\033[002mdef',
            'abc\033[40mdef',
            'abc\033[040mdef',
            'abc\033[0;1mdef',
            'abc\033[40;50mdef',
            'abc\033[50;30;40mdef',
            'abc\033[Adef',
            'abc\033[0Gdef',
            'abc\033[1;20;128Hdef',
        ]
        for datum in data:
            stream.wrapped.write.reset_mock()
            stream.write_and_convert( datum )
            self.assertEqual(
               [args[0] for args in stream.wrapped.write.call_args_list],
               [ ('abc',), ('def',) ]
            )

    def testWriteAndConvertSkipsEmptySnippets(self):
        stream = AnsiToWin32(Mock())
        stream.call_win32 = Mock()
        stream.write_and_convert( '\033[40m\033[41m' )
        self.assertFalse( stream.wrapped.write.called )

    def testWriteAndConvertCallsWin32WithParamsAndCommand(self):
        stream = AnsiToWin32(Mock())
        stream.convert = True
        stream.call_win32 = Mock()
        stream.extract_params = Mock(return_value='params')
        data = {
            'abc\033[adef':         ('a', 'params'),
            'abc\033[;;bdef':       ('b', 'params'),
            'abc\033[0cdef':        ('c', 'params'),
            'abc\033[;;0;;Gdef':    ('G', 'params'),
            'abc\033[1;20;128Hdef': ('H', 'params'),
        }
        for datum, expected in data.items():
            stream.call_win32.reset_mock()
            stream.write_and_convert( datum )
            self.assertEqual( stream.call_win32.call_args[0], expected )

    def test_reset_all_shouldnt_raise_on_closed_orig_stdout(self):
        stream = StringIO()
        converter = AnsiToWin32(stream)
        stream.close()

        converter.reset_all()

    def test_wrap_shouldnt_raise_on_closed_orig_stdout(self):
        stream = StringIO()
        stream.close()
        with \
            patch("colorama.ansitowin32.os.name", "nt"), \
            patch("colorama.ansitowin32.winapi_test", lambda: True):
                converter = AnsiToWin32(stream)
        self.assertTrue(converter.strip)
        self.assertFalse(converter.convert)

    def test_wrap_shouldnt_raise_on_missing_closed_attr(self):
        with \
            patch("colorama.ansitowin32.os.name", "nt"), \
            patch("colorama.ansitowin32.winapi_test", lambda: True):
                converter = AnsiToWin32(object())
        self.assertTrue(converter.strip)
        self.assertFalse(converter.convert)

    def testExtractParams(self):
        stream = AnsiToWin32(Mock())
        data = {
            '':               (0,),
            ';;':             (0,),
            '2':              (2,),
            ';;002;;':        (2,),
            '0;1':            (0, 1),
            ';;003;;456;;':   (3, 456),
            '11;22;33;44;55': (11, 22, 33, 44, 55),
        }
        for datum, expected in data.items():
            self.assertEqual(stream.extract_params('m', datum), expected)

    def testCallWin32UsesLookup(self):
        listener = Mock()
        stream = AnsiToWin32(listener)
        stream.win32_calls = {
            1: (lambda *_, **__: listener(11),),
            2: (lambda *_, **__: listener(22),),
            3: (lambda *_, **__: listener(33),),
        }
        stream.call_win32('m', (3, 1, 99, 2))
        self.assertEqual(
            [a[0][0] for a in listener.call_args_list],
            [33, 11, 22] )

    def test_osc_codes(self):
        mockStdout = Mock()
        stream = AnsiToWin32(mockStdout, convert=True)
        with patch('colorama.ansitowin32.winterm') as winterm:
            data = [
                '\033]0\x07',                      # missing arguments
                '\033]0;foo\x08',                  # wrong OSC command
                '\033]0;colorama_test_title\x07',  # should work
                '\033]1;colorama_test_title\x07',  # wrong set command
                '\033]2;colorama_test_title\x07',  # should work
                '\033]' + ';' * 64 + '\x08',       # see issue #247
            ]
            for code in data:
                stream.write(code)
            self.assertEqual(winterm.set_title.call_count, 2)

    def test_native_windows_ansi(self):
        with ExitStack() as stack:
            def p(a, b):
                stack.enter_context(patch(a, b, create=True))
            # Pretend to be on Windows
            p("colorama.ansitowin32.os.name", "nt")
            p("colorama.ansitowin32.winapi_test", lambda: True)
            p("colorama.win32.winapi_test", lambda: True)
            p("colorama.winterm.win32.windll", "non-None")
            p("colorama.winterm.get_osfhandle", lambda _: 1234)

            # Pretend that our mock stream has native ANSI support
            p(
                "colorama.winterm.win32.GetConsoleMode",
                lambda _: ENABLE_VIRTUAL_TERMINAL_PROCESSING,
            )
            SetConsoleMode = Mock()
            p("colorama.winterm.win32.SetConsoleMode", SetConsoleMode)

            stdout = Mock()
            stdout.closed = False
            stdout.isatty.return_value = True
            stdout.fileno.return_value = 1

            # Our fake console says it has native vt support, so AnsiToWin32 should
            # enable that support and do nothing else.
            stream = AnsiToWin32(stdout)
            SetConsoleMode.assert_called_with(1234, ENABLE_VIRTUAL_TERMINAL_PROCESSING)
            self.assertFalse(stream.strip)
            self.assertFalse(stream.convert)
            self.assertFalse(stream.should_wrap())

            # Now let's pretend we're on an old Windows console, that doesn't have
            # native ANSI support.
            p("colorama.winterm.win32.GetConsoleMode", lambda _: 0)
            SetConsoleMode = Mock()
            p("colorama.winterm.win32.SetConsoleMode", SetConsoleMode)

            stream = AnsiToWin32(stdout)
            SetConsoleMode.assert_called_with(1234, ENABLE_VIRTUAL_TERMINAL_PROCESSING)
            self.assertTrue(stream.strip)
            self.assertTrue(stream.convert)
            self.assertTrue(stream.should_wrap())


if __name__ == '__main__':
    main()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\groupby\transform\test_numba.py ===
import numpy as np
import pytest

from pandas.compat import is_platform_arm
from pandas.errors import NumbaUtilError

from pandas import (
    DataFrame,
    Series,
    option_context,
)
import pandas._testing as tm
from pandas.util.version import Version

pytestmark = [pytest.mark.single_cpu]

numba = pytest.importorskip("numba")
pytestmark.append(
    pytest.mark.skipif(
        Version(numba.__version__) == Version("0.61") and is_platform_arm(),
        reason=f"Segfaults on ARM platforms with numba {numba.__version__}",
    )
)


def test_correct_function_signature():
    pytest.importorskip("numba")

    def incorrect_function(x):
        return x + 1

    data = DataFrame(
        {"key": ["a", "a", "b", "b", "a"], "data": [1.0, 2.0, 3.0, 4.0, 5.0]},
        columns=["key", "data"],
    )
    with pytest.raises(NumbaUtilError, match="The first 2"):
        data.groupby("key").transform(incorrect_function, engine="numba")

    with pytest.raises(NumbaUtilError, match="The first 2"):
        data.groupby("key")["data"].transform(incorrect_function, engine="numba")


def test_check_nopython_kwargs():
    pytest.importorskip("numba")

    def incorrect_function(values, index):
        return values + 1

    data = DataFrame(
        {"key": ["a", "a", "b", "b", "a"], "data": [1.0, 2.0, 3.0, 4.0, 5.0]},
        columns=["key", "data"],
    )
    with pytest.raises(NumbaUtilError, match="numba does not support"):
        data.groupby("key").transform(incorrect_function, engine="numba", a=1)

    with pytest.raises(NumbaUtilError, match="numba does not support"):
        data.groupby("key")["data"].transform(incorrect_function, engine="numba", a=1)


@pytest.mark.filterwarnings("ignore")
# Filter warnings when parallel=True and the function can't be parallelized by Numba
@pytest.mark.parametrize("jit", [True, False])
@pytest.mark.parametrize("pandas_obj", ["Series", "DataFrame"])
@pytest.mark.parametrize("as_index", [True, False])
def test_numba_vs_cython(jit, pandas_obj, nogil, parallel, nopython, as_index):
    pytest.importorskip("numba")

    def func(values, index):
        return values + 1

    if jit:
        # Test accepted jitted functions
        import numba

        func = numba.jit(func)

    data = DataFrame(
        {0: ["a", "a", "b", "b", "a"], 1: [1.0, 2.0, 3.0, 4.0, 5.0]}, columns=[0, 1]
    )
    engine_kwargs = {"nogil": nogil, "parallel": parallel, "nopython": nopython}
    grouped = data.groupby(0, as_index=as_index)
    if pandas_obj == "Series":
        grouped = grouped[1]

    result = grouped.transform(func, engine="numba", engine_kwargs=engine_kwargs)
    expected = grouped.transform(lambda x: x + 1, engine="cython")

    tm.assert_equal(result, expected)


@pytest.mark.filterwarnings("ignore")
# Filter warnings when parallel=True and the function can't be parallelized by Numba
@pytest.mark.parametrize("jit", [True, False])
@pytest.mark.parametrize("pandas_obj", ["Series", "DataFrame"])
def test_cache(jit, pandas_obj, nogil, parallel, nopython):
    # Test that the functions are cached correctly if we switch functions
    pytest.importorskip("numba")

    def func_1(values, index):
        return values + 1

    def func_2(values, index):
        return values * 5

    if jit:
        import numba

        func_1 = numba.jit(func_1)
        func_2 = numba.jit(func_2)

    data = DataFrame(
        {0: ["a", "a", "b", "b", "a"], 1: [1.0, 2.0, 3.0, 4.0, 5.0]}, columns=[0, 1]
    )
    engine_kwargs = {"nogil": nogil, "parallel": parallel, "nopython": nopython}
    grouped = data.groupby(0)
    if pandas_obj == "Series":
        grouped = grouped[1]

    result = grouped.transform(func_1, engine="numba", engine_kwargs=engine_kwargs)
    expected = grouped.transform(lambda x: x + 1, engine="cython")
    tm.assert_equal(result, expected)

    result = grouped.transform(func_2, engine="numba", engine_kwargs=engine_kwargs)
    expected = grouped.transform(lambda x: x * 5, engine="cython")
    tm.assert_equal(result, expected)

    # Retest func_1 which should use the cache
    result = grouped.transform(func_1, engine="numba", engine_kwargs=engine_kwargs)
    expected = grouped.transform(lambda x: x + 1, engine="cython")
    tm.assert_equal(result, expected)


def test_use_global_config():
    pytest.importorskip("numba")

    def func_1(values, index):
        return values + 1

    data = DataFrame(
        {0: ["a", "a", "b", "b", "a"], 1: [1.0, 2.0, 3.0, 4.0, 5.0]}, columns=[0, 1]
    )
    grouped = data.groupby(0)
    expected = grouped.transform(func_1, engine="numba")
    with option_context("compute.use_numba", True):
        result = grouped.transform(func_1, engine=None)
    tm.assert_frame_equal(expected, result)


# TODO: Test more than just reductions (e.g. actually test transformations once we have
@pytest.mark.parametrize(
    "agg_func", [["min", "max"], "min", {"B": ["min", "max"], "C": "sum"}]
)
def test_string_cython_vs_numba(agg_func, numba_supported_reductions):
    pytest.importorskip("numba")
    agg_func, kwargs = numba_supported_reductions
    data = DataFrame(
        {0: ["a", "a", "b", "b", "a"], 1: [1.0, 2.0, 3.0, 4.0, 5.0]}, columns=[0, 1]
    )
    grouped = data.groupby(0)

    result = grouped.transform(agg_func, engine="numba", **kwargs)
    expected = grouped.transform(agg_func, engine="cython", **kwargs)
    tm.assert_frame_equal(result, expected)

    result = grouped[1].transform(agg_func, engine="numba", **kwargs)
    expected = grouped[1].transform(agg_func, engine="cython", **kwargs)
    tm.assert_series_equal(result, expected)


def test_args_not_cached():
    # GH 41647
    pytest.importorskip("numba")

    def sum_last(values, index, n):
        return values[-n:].sum()

    df = DataFrame({"id": [0, 0, 1, 1], "x": [1, 1, 1, 1]})
    grouped_x = df.groupby("id")["x"]
    result = grouped_x.transform(sum_last, 1, engine="numba")
    expected = Series([1.0] * 4, name="x")
    tm.assert_series_equal(result, expected)

    result = grouped_x.transform(sum_last, 2, engine="numba")
    expected = Series([2.0] * 4, name="x")
    tm.assert_series_equal(result, expected)


def test_index_data_correctly_passed():
    # GH 43133
    pytest.importorskip("numba")

    def f(values, index):
        return index - 1

    df = DataFrame({"group": ["A", "A", "B"], "v": [4, 5, 6]}, index=[-1, -2, -3])
    result = df.groupby("group").transform(f, engine="numba")
    expected = DataFrame([-4.0, -3.0, -2.0], columns=["v"], index=[-1, -2, -3])
    tm.assert_frame_equal(result, expected)


def test_engine_kwargs_not_cached():
    # If the user passes a different set of engine_kwargs don't return the same
    # jitted function
    pytest.importorskip("numba")
    nogil = True
    parallel = False
    nopython = True

    def func_kwargs(values, index):
        return nogil + parallel + nopython

    engine_kwargs = {"nopython": nopython, "nogil": nogil, "parallel": parallel}
    df = DataFrame({"value": [0, 0, 0]})
    result = df.groupby(level=0).transform(
        func_kwargs, engine="numba", engine_kwargs=engine_kwargs
    )
    expected = DataFrame({"value": [2.0, 2.0, 2.0]})
    tm.assert_frame_equal(result, expected)

    nogil = False
    engine_kwargs = {"nopython": nopython, "nogil": nogil, "parallel": parallel}
    result = df.groupby(level=0).transform(
        func_kwargs, engine="numba", engine_kwargs=engine_kwargs
    )
    expected = DataFrame({"value": [1.0, 1.0, 1.0]})
    tm.assert_frame_equal(result, expected)


@pytest.mark.filterwarnings("ignore")
def test_multiindex_one_key(nogil, parallel, nopython):
    pytest.importorskip("numba")

    def numba_func(values, index):
        return 1

    df = DataFrame([{"A": 1, "B": 2, "C": 3}]).set_index(["A", "B"])
    engine_kwargs = {"nopython": nopython, "nogil": nogil, "parallel": parallel}
    result = df.groupby("A").transform(
        numba_func, engine="numba", engine_kwargs=engine_kwargs
    )
    expected = DataFrame([{"A": 1, "B": 2, "C": 1.0}]).set_index(["A", "B"])
    tm.assert_frame_equal(result, expected)


def test_multiindex_multi_key_not_supported(nogil, parallel, nopython):
    pytest.importorskip("numba")

    def numba_func(values, index):
        return 1

    df = DataFrame([{"A": 1, "B": 2, "C": 3}]).set_index(["A", "B"])
    engine_kwargs = {"nopython": nopython, "nogil": nogil, "parallel": parallel}
    with pytest.raises(NotImplementedError, match="more than 1 grouping labels"):
        df.groupby(["A", "B"]).transform(
            numba_func, engine="numba", engine_kwargs=engine_kwargs
        )


def test_multilabel_numba_vs_cython(numba_supported_reductions):
    pytest.importorskip("numba")
    reduction, kwargs = numba_supported_reductions
    df = DataFrame(
        {
            "A": ["foo", "bar", "foo", "bar", "foo", "bar", "foo", "foo"],
            "B": ["one", "one", "two", "three", "two", "two", "one", "three"],
            "C": np.random.default_rng(2).standard_normal(8),
            "D": np.random.default_rng(2).standard_normal(8),
        }
    )
    gb = df.groupby(["A", "B"])
    res_agg = gb.transform(reduction, engine="numba", **kwargs)
    expected_agg = gb.transform(reduction, engine="cython", **kwargs)
    tm.assert_frame_equal(res_agg, expected_agg)


def test_multilabel_udf_numba_vs_cython():
    pytest.importorskip("numba")
    df = DataFrame(
        {
            "A": ["foo", "bar", "foo", "bar", "foo", "bar", "foo", "foo"],
            "B": ["one", "one", "two", "three", "two", "two", "one", "three"],
            "C": np.random.default_rng(2).standard_normal(8),
            "D": np.random.default_rng(2).standard_normal(8),
        }
    )
    gb = df.groupby(["A", "B"])
    result = gb.transform(
        lambda values, index: (values - values.min()) / (values.max() - values.min()),
        engine="numba",
    )
    expected = gb.transform(
        lambda x: (x - x.min()) / (x.max() - x.min()), engine="cython"
    )
    tm.assert_frame_equal(result, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\tests\test_multivariate_resultants.py ===
"""Tests for Dixon's and Macaulay's classes. """

from sympy.matrices.dense import Matrix
from sympy.polys.polytools import factor
from sympy.core import symbols
from sympy.tensor.indexed import IndexedBase

from sympy.polys.multivariate_resultants import (DixonResultant,
                                                 MacaulayResultant)

c, d = symbols("a, b")
x, y = symbols("x, y")

p =  c * x + y
q =  x + d * y

dixon = DixonResultant(polynomials=[p, q], variables=[x, y])
macaulay = MacaulayResultant(polynomials=[p, q], variables=[x, y])

def test_dixon_resultant_init():
    """Test init method of DixonResultant."""
    a = IndexedBase("alpha")

    assert dixon.polynomials == [p, q]
    assert dixon.variables == [x, y]
    assert dixon.n == 2
    assert dixon.m == 2
    assert dixon.dummy_variables == [a[0], a[1]]

def test_get_dixon_polynomial_numerical():
    """Test Dixon's polynomial for a numerical example."""
    a = IndexedBase("alpha")

    p = x + y
    q = x ** 2 + y **3
    h = x ** 2 + y

    dixon = DixonResultant([p, q, h], [x, y])
    polynomial = -x * y ** 2 * a[0] - x * y ** 2 * a[1] - x * y * a[0] \
    * a[1] - x * y * a[1] ** 2 - x * a[0] * a[1] ** 2 + x * a[0] - \
    y ** 2 * a[0] * a[1] + y ** 2 * a[1] - y * a[0] * a[1] ** 2 + y * \
    a[1] ** 2

    assert dixon.get_dixon_polynomial().as_expr().expand() == polynomial

def test_get_max_degrees():
    """Tests max degrees function."""

    p = x + y
    q = x ** 2 + y **3
    h = x ** 2 + y

    dixon = DixonResultant(polynomials=[p, q, h], variables=[x, y])
    dixon_polynomial = dixon.get_dixon_polynomial()

    assert dixon.get_max_degrees(dixon_polynomial) == [1, 2]

def test_get_dixon_matrix():
    """Test Dixon's resultant for a numerical example."""

    x, y = symbols('x, y')

    p = x + y
    q = x ** 2 + y ** 3
    h = x ** 2 + y

    dixon = DixonResultant([p, q, h], [x, y])
    polynomial = dixon.get_dixon_polynomial()

    assert dixon.get_dixon_matrix(polynomial).det() == 0

def test_get_dixon_matrix_example_two():
    """Test Dixon's matrix for example from [Palancz08]_."""
    x, y, z = symbols('x, y, z')

    f = x ** 2 + y ** 2 - 1 + z * 0
    g = x ** 2 + z ** 2 - 1 + y * 0
    h = y ** 2 + z ** 2 - 1

    example_two = DixonResultant([f, g, h], [y, z])
    poly = example_two.get_dixon_polynomial()
    matrix = example_two.get_dixon_matrix(poly)

    expr = 1 - 8 * x ** 2 + 24 * x ** 4 - 32 * x ** 6 + 16 * x ** 8
    assert (matrix.det() - expr).expand() == 0

def test_KSY_precondition():
    """Tests precondition for KSY Resultant."""
    A, B, C = symbols('A, B, C')

    m1 = Matrix([[1, 2, 3],
                 [4, 5, 12],
                 [6, 7, 18]])

    m2 = Matrix([[0, C**2],
                 [-2 * C, -C ** 2]])

    m3 = Matrix([[1, 0],
                 [0, 1]])

    m4 = Matrix([[A**2, 0, 1],
                 [A, 1, 1 / A]])

    m5 = Matrix([[5, 1],
                 [2, B],
                 [0, 1],
                 [0, 0]])

    assert dixon.KSY_precondition(m1) == False
    assert dixon.KSY_precondition(m2) == True
    assert dixon.KSY_precondition(m3) == True
    assert dixon.KSY_precondition(m4) == False
    assert dixon.KSY_precondition(m5) == True

def test_delete_zero_rows_and_columns():
    """Tests method for deleting rows and columns containing only zeros."""
    A, B, C = symbols('A, B, C')

    m1 = Matrix([[0, 0],
                 [0, 0],
                 [1, 2]])

    m2 = Matrix([[0, 1, 2],
                 [0, 3, 4],
                 [0, 5, 6]])

    m3 = Matrix([[0, 0, 0, 0],
                 [0, 1, 2, 0],
                 [0, 3, 4, 0],
                 [0, 0, 0, 0]])

    m4 = Matrix([[1, 0, 2],
                 [0, 0, 0],
                 [3, 0, 4]])

    m5 = Matrix([[0, 0, 0, 1],
                 [0, 0, 0, 2],
                 [0, 0, 0, 3],
                 [0, 0, 0, 4]])

    m6 = Matrix([[0, 0, A],
                 [B, 0, 0],
                 [0, 0, C]])

    assert dixon.delete_zero_rows_and_columns(m1) == Matrix([[1, 2]])

    assert dixon.delete_zero_rows_and_columns(m2) == Matrix([[1, 2],
                                                             [3, 4],
                                                             [5, 6]])

    assert dixon.delete_zero_rows_and_columns(m3) == Matrix([[1, 2],
                                                             [3, 4]])

    assert dixon.delete_zero_rows_and_columns(m4) == Matrix([[1, 2],
                                                             [3, 4]])

    assert dixon.delete_zero_rows_and_columns(m5) == Matrix([[1],
                                                             [2],
                                                             [3],
                                                             [4]])

    assert dixon.delete_zero_rows_and_columns(m6) == Matrix([[0, A],
                                                             [B, 0],
                                                             [0, C]])

def test_product_leading_entries():
    """Tests product of leading entries method."""
    A, B = symbols('A, B')

    m1 = Matrix([[1, 2, 3],
                 [0, 4, 5],
                 [0, 0, 6]])

    m2 = Matrix([[0, 0, 1],
                 [2, 0, 3]])

    m3 = Matrix([[0, 0, 0],
                 [1, 2, 3],
                 [0, 0, 0]])

    m4 = Matrix([[0, 0, A],
                 [1, 2, 3],
                 [B, 0, 0]])

    assert dixon.product_leading_entries(m1) == 24
    assert dixon.product_leading_entries(m2) == 2
    assert dixon.product_leading_entries(m3) == 1
    assert dixon.product_leading_entries(m4) == A * B

def test_get_KSY_Dixon_resultant_example_one():
    """Tests the KSY Dixon resultant for example one"""
    x, y, z = symbols('x, y, z')

    p = x * y * z
    q = x**2 - z**2
    h = x + y + z
    dixon = DixonResultant([p, q, h], [x, y])
    dixon_poly = dixon.get_dixon_polynomial()
    dixon_matrix = dixon.get_dixon_matrix(dixon_poly)
    D = dixon.get_KSY_Dixon_resultant(dixon_matrix)

    assert D == -z**3

def test_get_KSY_Dixon_resultant_example_two():
    """Tests the KSY Dixon resultant for example two"""
    x, y, A = symbols('x, y, A')

    p = x * y + x * A + x - A**2 - A + y**2 + y
    q = x**2 + x * A - x + x * y + y * A - y
    h = x**2 + x * y + 2 * x - x * A - y * A - 2 * A

    dixon = DixonResultant([p, q, h], [x, y])
    dixon_poly = dixon.get_dixon_polynomial()
    dixon_matrix = dixon.get_dixon_matrix(dixon_poly)
    D = factor(dixon.get_KSY_Dixon_resultant(dixon_matrix))

    assert D == -8*A*(A - 1)*(A + 2)*(2*A - 1)**2

def test_macaulay_resultant_init():
    """Test init method of MacaulayResultant."""

    assert macaulay.polynomials == [p, q]
    assert macaulay.variables == [x, y]
    assert macaulay.n == 2
    assert macaulay.degrees == [1, 1]
    assert macaulay.degree_m == 1
    assert macaulay.monomials_size == 2

def test_get_degree_m():
    assert macaulay._get_degree_m() == 1

def test_get_size():
    assert macaulay.get_size() == 2

def test_macaulay_example_one():
    """Tests the Macaulay for example from [Bruce97]_"""

    x, y, z = symbols('x, y, z')
    a_1_1, a_1_2, a_1_3 = symbols('a_1_1, a_1_2, a_1_3')
    a_2_2, a_2_3, a_3_3 = symbols('a_2_2, a_2_3, a_3_3')
    b_1_1, b_1_2, b_1_3 = symbols('b_1_1, b_1_2, b_1_3')
    b_2_2, b_2_3, b_3_3 = symbols('b_2_2, b_2_3, b_3_3')
    c_1, c_2, c_3 = symbols('c_1, c_2, c_3')

    f_1 = a_1_1 * x ** 2 + a_1_2 * x * y + a_1_3 * x * z + \
          a_2_2 * y ** 2 + a_2_3 * y * z + a_3_3 * z ** 2
    f_2 = b_1_1 * x ** 2 + b_1_2 * x * y + b_1_3 * x * z + \
          b_2_2 * y ** 2 + b_2_3 * y * z + b_3_3 * z ** 2
    f_3 = c_1 * x + c_2 * y + c_3 * z

    mac = MacaulayResultant([f_1, f_2, f_3], [x, y, z])

    assert mac.degrees == [2, 2, 1]
    assert mac.degree_m == 3

    assert mac.monomial_set == [x ** 3, x ** 2 * y, x ** 2 * z,
                                x * y ** 2,
                                x * y * z, x * z ** 2, y ** 3,
                                y ** 2 *z, y * z ** 2, z ** 3]
    assert mac.monomials_size == 10
    assert mac.get_row_coefficients() == [[x, y, z], [x, y, z],
                                          [x * y, x * z, y * z, z ** 2]]

    matrix = mac.get_matrix()
    assert matrix.shape == (mac.monomials_size, mac.monomials_size)
    assert mac.get_submatrix(matrix) == Matrix([[a_1_1, a_2_2],
                                                [b_1_1, b_2_2]])

def test_macaulay_example_two():
    """Tests the Macaulay formulation for example from [Stiller96]_."""

    x, y, z = symbols('x, y, z')
    a_0, a_1, a_2 = symbols('a_0, a_1, a_2')
    b_0, b_1, b_2 = symbols('b_0, b_1, b_2')
    c_0, c_1, c_2, c_3, c_4 = symbols('c_0, c_1, c_2, c_3, c_4')

    f = a_0 * y -  a_1 * x + a_2 * z
    g = b_1 * x ** 2 + b_0 * y ** 2 - b_2 * z ** 2
    h = c_0 * y - c_1 * x ** 3 + c_2 * x ** 2 * z - c_3 * x * z ** 2 + \
        c_4 * z ** 3

    mac = MacaulayResultant([f, g, h], [x, y, z])

    assert mac.degrees == [1, 2, 3]
    assert mac.degree_m == 4
    assert mac.monomials_size == 15
    assert len(mac.get_row_coefficients()) == mac.n

    matrix = mac.get_matrix()
    assert matrix.shape == (mac.monomials_size, mac.monomials_size)
    assert mac.get_submatrix(matrix) == Matrix([[-a_1, a_0, a_2, 0],
                                                [0, -a_1, 0, 0],
                                                [0, 0, -a_1, 0],
                                                [0, 0, 0, -a_1]])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\sets\tests\test_conditionset.py ===
from sympy.core.expr import unchanged
from sympy.sets import (ConditionSet, Intersection, FiniteSet,
    EmptySet, Union, Contains, ImageSet)
from sympy.sets.sets import SetKind
from sympy.core.function import (Function, Lambda)
from sympy.core.mod import Mod
from sympy.core.kind import NumberKind
from sympy.core.numbers import (oo, pi)
from sympy.core.relational import (Eq, Ne)
from sympy.core.singleton import S
from sympy.core.symbol import (Symbol, symbols)
from sympy.functions.elementary.complexes import Abs
from sympy.functions.elementary.trigonometric import (asin, sin)
from sympy.logic.boolalg import And
from sympy.matrices.dense import Matrix
from sympy.matrices.expressions.matexpr import MatrixSymbol
from sympy.sets.sets import Interval
from sympy.testing.pytest import raises, warns_deprecated_sympy


w = Symbol('w')
x = Symbol('x')
y = Symbol('y')
z = Symbol('z')
f = Function('f')


def test_CondSet():
    sin_sols_principal = ConditionSet(x, Eq(sin(x), 0),
                                      Interval(0, 2*pi, False, True))
    assert pi in sin_sols_principal
    assert pi/2 not in sin_sols_principal
    assert 3*pi not in sin_sols_principal
    assert oo not in sin_sols_principal
    assert 5 in ConditionSet(x, x**2 > 4, S.Reals)
    assert 1 not in ConditionSet(x, x**2 > 4, S.Reals)
    # in this case, 0 is not part of the base set so
    # it can't be in any subset selected by the condition
    assert 0 not in ConditionSet(x, y > 5, Interval(1, 7))
    # since 'in' requires a true/false, the following raises
    # an error because the given value provides no information
    # for the condition to evaluate (since the condition does
    # not depend on the dummy symbol): the result is `y > 5`.
    # In this case, ConditionSet is just acting like
    # Piecewise((Interval(1, 7), y > 5), (S.EmptySet, True)).
    raises(TypeError, lambda: 6 in ConditionSet(x, y > 5,
        Interval(1, 7)))

    X = MatrixSymbol('X', 2, 2)
    matrix_set = ConditionSet(X, Eq(X*Matrix([[1, 1], [1, 1]]), X))
    Y = Matrix([[0, 0], [0, 0]])
    assert matrix_set.contains(Y).doit() is S.true
    Z = Matrix([[1, 2], [3, 4]])
    assert matrix_set.contains(Z).doit() is S.false

    assert isinstance(ConditionSet(x, x < 1, {x, y}).base_set,
        FiniteSet)
    raises(TypeError, lambda: ConditionSet(x, x + 1, {x, y}))
    raises(TypeError, lambda: ConditionSet(x, x, 1))

    I = S.Integers
    U = S.UniversalSet
    C = ConditionSet
    assert C(x, False, I) is S.EmptySet
    assert C(x, True, I) is I
    assert C(x, x < 1, C(x, x < 2, I)
        ) == C(x, (x < 1) & (x < 2), I)
    assert C(y, y < 1, C(x, y < 2, I)
        ) == C(x, (x < 1) & (y < 2), I), C(y, y < 1, C(x, y < 2, I))
    assert C(y, y < 1, C(x, x < 2, I)
        ) == C(y, (y < 1) & (y < 2), I)
    assert C(y, y < 1, C(x, y < x, I)
        ) == C(x, (x < 1) & (y < x), I)
    assert unchanged(C, y, x < 1, C(x, y < x, I))
    assert ConditionSet(x, x < 1).base_set is U
    # arg checking is not done at instantiation but this
    # will raise an error when containment is tested
    assert ConditionSet((x,), x < 1).base_set is U

    c = ConditionSet((x, y), x < y, I**2)
    assert (1, 2) in c
    assert (1, pi) not in c

    raises(TypeError, lambda: C(x, x > 1, C((x, y), x > 1, I**2)))
    # signature mismatch since only 3 args are accepted
    raises(TypeError, lambda: C((x, y), x + y < 2, U, U))


def test_CondSet_intersect():
    input_conditionset = ConditionSet(x, x**2 > 4, Interval(1, 4, False,
        False))
    other_domain = Interval(0, 3, False, False)
    output_conditionset = ConditionSet(x, x**2 > 4, Interval(
        1, 3, False, False))
    assert Intersection(input_conditionset, other_domain
        ) == output_conditionset


def test_issue_9849():
    assert ConditionSet(x, Eq(x, x), S.Naturals
        ) is S.Naturals
    assert ConditionSet(x, Eq(Abs(sin(x)), -1), S.Naturals
        ) == S.EmptySet


def test_simplified_FiniteSet_in_CondSet():
    assert ConditionSet(x, And(x < 1, x > -3), FiniteSet(0, 1, 2)
        ) == FiniteSet(0)
    assert ConditionSet(x, x < 0, FiniteSet(0, 1, 2)) == EmptySet
    assert ConditionSet(x, And(x < -3), EmptySet) == EmptySet
    y = Symbol('y')
    assert (ConditionSet(x, And(x > 0), FiniteSet(-1, 0, 1, y)) ==
        Union(FiniteSet(1), ConditionSet(x, And(x > 0), FiniteSet(y))))
    assert (ConditionSet(x, Eq(Mod(x, 3), 1), FiniteSet(1, 4, 2, y)) ==
        Union(FiniteSet(1, 4), ConditionSet(x, Eq(Mod(x, 3), 1),
        FiniteSet(y))))


def test_free_symbols():
    assert ConditionSet(x, Eq(y, 0), FiniteSet(z)
        ).free_symbols == {y, z}
    assert ConditionSet(x, Eq(x, 0), FiniteSet(z)
        ).free_symbols == {z}
    assert ConditionSet(x, Eq(x, 0), FiniteSet(x, z)
        ).free_symbols == {x, z}
    assert ConditionSet(x, Eq(x, 0), ImageSet(Lambda(y, y**2),
        S.Integers)).free_symbols == set()


def test_bound_symbols():
    assert ConditionSet(x, Eq(y, 0), FiniteSet(z)
        ).bound_symbols == [x]
    assert ConditionSet(x, Eq(x, 0), FiniteSet(x, y)
        ).bound_symbols == [x]
    assert ConditionSet(x, x < 10, ImageSet(Lambda(y, y**2), S.Integers)
        ).bound_symbols == [x]
    assert ConditionSet(x, x < 10, ConditionSet(y, y > 1, S.Integers)
        ).bound_symbols == [x]


def test_as_dummy():
    _0, _1 = symbols('_0 _1')
    assert ConditionSet(x, x < 1, Interval(y, oo)
        ).as_dummy() == ConditionSet(_0, _0 < 1, Interval(y, oo))
    assert ConditionSet(x, x < 1, Interval(x, oo)
        ).as_dummy() == ConditionSet(_0, _0 < 1, Interval(x, oo))
    assert ConditionSet(x, x < 1, ImageSet(Lambda(y, y**2), S.Integers)
        ).as_dummy() == ConditionSet(
            _0, _0 < 1, ImageSet(Lambda(_0, _0**2), S.Integers))
    e = ConditionSet((x, y), x <= y, S.Reals**2)
    assert e.bound_symbols == [x, y]
    assert e.as_dummy() == ConditionSet((_0, _1), _0 <= _1, S.Reals**2)
    assert e.as_dummy() == ConditionSet((y, x), y <= x, S.Reals**2
        ).as_dummy()


def test_subs_CondSet():
    s = FiniteSet(z, y)
    c = ConditionSet(x, x < 2, s)
    assert c.subs(x, y) == c
    assert c.subs(z, y) == ConditionSet(x, x < 2, FiniteSet(y))
    assert c.xreplace({x: y}) == ConditionSet(y, y < 2, s)

    assert ConditionSet(x, x < y, s
        ).subs(y, w) == ConditionSet(x, x < w, s.subs(y, w))
    # if the user uses assumptions that cause the condition
    # to evaluate, that can't be helped from SymPy's end
    n = Symbol('n', negative=True)
    assert ConditionSet(n, 0 < n, S.Integers) is S.EmptySet
    p = Symbol('p', positive=True)
    assert ConditionSet(n, n < y, S.Integers
        ).subs(n, x) == ConditionSet(n, n < y, S.Integers)
    raises(ValueError, lambda: ConditionSet(
        x + 1, x < 1, S.Integers))
    assert ConditionSet(
        p, n < x, Interval(-5, 5)).subs(x, p) == Interval(-5, 5), ConditionSet(
        p, n < x, Interval(-5, 5)).subs(x, p)
    assert ConditionSet(
        n, n < x, Interval(-oo, 0)).subs(x, p
        ) == Interval(-oo, 0)

    assert ConditionSet(f(x), f(x) < 1, {w, z}
        ).subs(f(x), y) == ConditionSet(f(x), f(x) < 1, {w, z})

    # issue 17341
    k = Symbol('k')
    img1 = ImageSet(Lambda(k, 2*k*pi + asin(y)), S.Integers)
    img2 = ImageSet(Lambda(k, 2*k*pi + asin(S.One/3)), S.Integers)
    assert ConditionSet(x, Contains(
        y, Interval(-1,1)), img1).subs(y, S.One/3).dummy_eq(img2)

    assert (0, 1) in ConditionSet((x, y), x + y < 3, S.Integers**2)

    raises(TypeError, lambda: ConditionSet(n, n < -10, Interval(0, 10)))


def test_subs_CondSet_tebr():
    with warns_deprecated_sympy():
        assert ConditionSet((x, y), {x + 1, x + y}, S.Reals**2) == \
            ConditionSet((x, y), Eq(x + 1, 0) & Eq(x + y, 0), S.Reals**2)


def test_dummy_eq():
    C = ConditionSet
    I = S.Integers
    c = C(x, x < 1, I)
    assert c.dummy_eq(C(y, y < 1, I))
    assert c.dummy_eq(1) == False
    assert c.dummy_eq(C(x, x < 1, S.Reals)) == False

    c1 = ConditionSet((x, y), Eq(x + 1, 0) & Eq(x + y, 0), S.Reals**2)
    c2 = ConditionSet((x, y), Eq(x + 1, 0) & Eq(x + y, 0), S.Reals**2)
    c3 = ConditionSet((x, y), Eq(x + 1, 0) & Eq(x + y, 0), S.Complexes**2)
    assert c1.dummy_eq(c2)
    assert c1.dummy_eq(c3) is False
    assert c.dummy_eq(c1) is False
    assert c1.dummy_eq(c) is False

    # issue 19496
    m = Symbol('m')
    n = Symbol('n')
    a = Symbol('a')
    d1 = ImageSet(Lambda(m, m*pi), S.Integers)
    d2 = ImageSet(Lambda(n, n*pi), S.Integers)
    c1 = ConditionSet(x, Ne(a, 0), d1)
    c2 = ConditionSet(x, Ne(a, 0), d2)
    assert c1.dummy_eq(c2)


def test_contains():
    assert 6 in ConditionSet(x, x > 5, Interval(1, 7))
    assert (8 in ConditionSet(x, y > 5, Interval(1, 7))) is False
    # `in` should give True or False; in this case there is not
    # enough information for that result
    raises(TypeError,
        lambda: 6 in ConditionSet(x, y > 5, Interval(1, 7)))
    # here, there is enough information but the comparison is
    # not defined
    raises(TypeError, lambda: 0 in ConditionSet(x, 1/x >= 0, S.Reals))
    assert ConditionSet(x, y > 5, Interval(1, 7)
        ).contains(6) == (y > 5)
    assert ConditionSet(x, y > 5, Interval(1, 7)
        ).contains(8) is S.false
    assert ConditionSet(x, y > 5, Interval(1, 7)
        ).contains(w) == And(Contains(w, Interval(1, 7)), y > 5)
    # This returns an unevaluated Contains object
    # because 1/0 should not be defined for 1 and 0 in the context of
    # reals.
    assert ConditionSet(x, 1/x >= 0, S.Reals).contains(0) == \
        Contains(0, ConditionSet(x, 1/x >= 0, S.Reals), evaluate=False)
    c = ConditionSet((x, y), x + y > 1, S.Integers**2)
    assert not c.contains(1)
    assert c.contains((2, 1))
    assert not c.contains((0, 1))
    c = ConditionSet((w, (x, y)), w + x + y > 1, S.Integers*S.Integers**2)
    assert not c.contains(1)
    assert not c.contains((1, 2))
    assert not c.contains(((1, 2), 3))
    assert not c.contains(((1, 2), (3, 4)))
    assert c.contains((1, (3, 4)))


def test_as_relational():
    assert ConditionSet((x, y), x > 1, S.Integers**2).as_relational((x, y)
        ) == (x > 1) & Contains(x, S.Integers) & Contains(y, S.Integers)
    assert ConditionSet(x, x > 1, S.Integers).as_relational(x
        ) == Contains(x, S.Integers) & (x > 1)


def test_flatten():
    """Tests whether there is basic denesting functionality"""
    inner = ConditionSet(x, sin(x) + x > 0)
    outer = ConditionSet(x, Contains(x, inner), S.Reals)
    assert outer == ConditionSet(x, sin(x) + x > 0, S.Reals)

    inner = ConditionSet(y, sin(y) + y > 0)
    outer = ConditionSet(x, Contains(y, inner), S.Reals)
    assert outer != ConditionSet(x, sin(x) + x > 0, S.Reals)

    inner = ConditionSet(x, sin(x) + x > 0).intersect(Interval(-1, 1))
    outer = ConditionSet(x, Contains(x, inner), S.Reals)
    assert outer == ConditionSet(x, sin(x) + x > 0, Interval(-1, 1))


def test_duplicate():
    from sympy.core.function import BadSignatureError
    # test coverage for line 95 in conditionset.py, check for duplicates in symbols
    dup = symbols('a,a')
    raises(BadSignatureError, lambda: ConditionSet(dup, x < 0))


def test_SetKind_ConditionSet():
    assert ConditionSet(x, Eq(sin(x), 0), Interval(0, 2*pi)).kind is SetKind(NumberKind)
    assert ConditionSet(x, x < 0).kind is SetKind(NumberKind)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_core\tests\test_multithreading.py ===
import concurrent.futures
import string
import threading

import pytest

import numpy as np
from numpy._core import _rational_tests
from numpy.testing import IS_64BIT, IS_WASM
from numpy.testing._private.utils import run_threaded

if IS_WASM:
    pytest.skip(allow_module_level=True, reason="no threading support in wasm")


def test_parallel_randomstate_creation():
    # if the coercion cache is enabled and not thread-safe, creating
    # RandomState instances simultaneously leads to a data race
    def func(seed):
        np.random.RandomState(seed)

    run_threaded(func, 500, pass_count=True)


def test_parallel_ufunc_execution():
    # if the loop data cache or dispatch cache are not thread-safe
    # computing ufuncs simultaneously in multiple threads leads
    # to a data race that causes crashes or spurious exceptions
    def func():
        arr = np.random.random((25,))
        np.isnan(arr)

    run_threaded(func, 500)

    # see gh-26690
    NUM_THREADS = 50

    a = np.ones(1000)

    def f(b):
        b.wait()
        return a.sum()

    run_threaded(f, NUM_THREADS, pass_barrier=True)


def test_temp_elision_thread_safety():
    amid = np.ones(50000)
    bmid = np.ones(50000)
    alarge = np.ones(1000000)
    blarge = np.ones(1000000)

    def func(count):
        if count % 4 == 0:
            (amid * 2) + bmid
        elif count % 4 == 1:
            (amid + bmid) - 2
        elif count % 4 == 2:
            (alarge * 2) + blarge
        else:
            (alarge + blarge) - 2

    run_threaded(func, 100, pass_count=True)


def test_eigvalsh_thread_safety():
    # if lapack isn't thread safe this will randomly segfault or error
    # see gh-24512
    rng = np.random.RandomState(873699172)
    matrices = (
        rng.random((5, 10, 10, 3, 3)),
        rng.random((5, 10, 10, 3, 3)),
    )

    run_threaded(lambda i: np.linalg.eigvalsh(matrices[i]), 2,
                 pass_count=True)


def test_printoptions_thread_safety():
    # until NumPy 2.1 the printoptions state was stored in globals
    # this verifies that they are now stored in a context variable
    b = threading.Barrier(2)

    def legacy_113():
        np.set_printoptions(legacy='1.13', precision=12)
        b.wait()
        po = np.get_printoptions()
        assert po['legacy'] == '1.13'
        assert po['precision'] == 12
        orig_linewidth = po['linewidth']
        with np.printoptions(linewidth=34, legacy='1.21'):
            po = np.get_printoptions()
            assert po['legacy'] == '1.21'
            assert po['precision'] == 12
            assert po['linewidth'] == 34
        po = np.get_printoptions()
        assert po['linewidth'] == orig_linewidth
        assert po['legacy'] == '1.13'
        assert po['precision'] == 12

    def legacy_125():
        np.set_printoptions(legacy='1.25', precision=7)
        b.wait()
        po = np.get_printoptions()
        assert po['legacy'] == '1.25'
        assert po['precision'] == 7
        orig_linewidth = po['linewidth']
        with np.printoptions(linewidth=6, legacy='1.13'):
            po = np.get_printoptions()
            assert po['legacy'] == '1.13'
            assert po['precision'] == 7
            assert po['linewidth'] == 6
        po = np.get_printoptions()
        assert po['linewidth'] == orig_linewidth
        assert po['legacy'] == '1.25'
        assert po['precision'] == 7

    task1 = threading.Thread(target=legacy_113)
    task2 = threading.Thread(target=legacy_125)

    task1.start()
    task2.start()


def test_parallel_reduction():
    # gh-28041
    NUM_THREADS = 50

    x = np.arange(1000)

    def closure(b):
        b.wait()
        np.sum(x)

    run_threaded(closure, NUM_THREADS, pass_barrier=True)


def test_parallel_flat_iterator():
    # gh-28042
    x = np.arange(20).reshape(5, 4).T

    def closure(b):
        b.wait()
        for _ in range(100):
            list(x.flat)

    run_threaded(closure, outer_iterations=100, pass_barrier=True)

    # gh-28143
    def prepare_args():
        return [np.arange(10)]

    def closure(x, b):
        b.wait()
        for _ in range(100):
            y = np.arange(10)
            y.flat[x] = x

    run_threaded(closure, pass_barrier=True, prepare_args=prepare_args)


def test_multithreaded_repeat():
    x0 = np.arange(10)

    def closure(b):
        b.wait()
        for _ in range(100):
            x = np.repeat(x0, 2, axis=0)[::2]

    run_threaded(closure, max_workers=10, pass_barrier=True)


def test_structured_advanced_indexing():
    # Test that copyswap(n) used by integer array indexing is threadsafe
    # for structured datatypes, see gh-15387. This test can behave randomly.

    # Create a deeply nested dtype to make a failure more likely:
    dt = np.dtype([("", "f8")])
    dt = np.dtype([("", dt)] * 2)
    dt = np.dtype([("", dt)] * 2)
    # The array should be large enough to likely run into threading issues
    arr = np.random.uniform(size=(6000, 8)).view(dt)[:, 0]

    rng = np.random.default_rng()

    def func(arr):
        indx = rng.integers(0, len(arr), size=6000, dtype=np.intp)
        arr[indx]

    tpe = concurrent.futures.ThreadPoolExecutor(max_workers=8)
    futures = [tpe.submit(func, arr) for _ in range(10)]
    for f in futures:
        f.result()

    assert arr.dtype is dt


def test_structured_threadsafety2():
    # Nonzero (and some other functions) should be threadsafe for
    # structured datatypes, see gh-15387. This test can behave randomly.
    from concurrent.futures import ThreadPoolExecutor

    # Create a deeply nested dtype to make a failure more likely:
    dt = np.dtype([("", "f8")])
    dt = np.dtype([("", dt)])
    dt = np.dtype([("", dt)] * 2)
    # The array should be large enough to likely run into threading issues
    arr = np.random.uniform(size=(5000, 4)).view(dt)[:, 0]

    def func(arr):
        arr.nonzero()

    tpe = ThreadPoolExecutor(max_workers=8)
    futures = [tpe.submit(func, arr) for _ in range(10)]
    for f in futures:
        f.result()

    assert arr.dtype is dt


def test_stringdtype_multithreaded_access_and_mutation(
        dtype, random_string_list):
    # this test uses an RNG and may crash or cause deadlocks if there is a
    # threading bug
    rng = np.random.default_rng(0x4D3D3D3)

    chars = list(string.ascii_letters + string.digits)
    chars = np.array(chars, dtype="U1")
    ret = rng.choice(chars, size=100 * 10, replace=True)
    random_string_list = ret.view("U100")

    def func(arr):
        rnd = rng.random()
        # either write to random locations in the array, compute a ufunc, or
        # re-initialize the array
        if rnd < 0.25:
            num = np.random.randint(0, arr.size)
            arr[num] = arr[num] + "hello"
        elif rnd < 0.5:
            if rnd < 0.375:
                np.add(arr, arr)
            else:
                np.add(arr, arr, out=arr)
        elif rnd < 0.75:
            if rnd < 0.875:
                np.multiply(arr, np.int64(2))
            else:
                np.multiply(arr, np.int64(2), out=arr)
        else:
            arr[:] = random_string_list

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as tpe:
        arr = np.array(random_string_list, dtype=dtype)
        futures = [tpe.submit(func, arr) for _ in range(500)]

        for f in futures:
            f.result()


@pytest.mark.skipif(
    not IS_64BIT,
    reason="Sometimes causes failures or crashes due to OOM on 32 bit runners"
)
def test_legacy_usertype_cast_init_thread_safety():
    def closure(b):
        b.wait()
        np.full((10, 10), 1, _rational_tests.rational)

    run_threaded(closure, 250, pass_barrier=True)

@pytest.mark.parametrize("dtype", [bool, int, float])
def test_nonzero(dtype):
    # See: gh-28361
    #
    # np.nonzero uses np.count_nonzero to determine the size of the output array
    # In a second pass the indices of the non-zero elements are determined, but they can have changed
    #
    # This test triggers a data race which is suppressed in the TSAN CI. The test is to ensure
    # np.nonzero does not generate a segmentation fault
    x = np.random.randint(4, size=100).astype(dtype)

    def func(index):
        for _ in range(10):
            if index == 0:
                x[::2] = np.random.randint(2)
            else:
                try:
                    _ = np.nonzero(x)
                except RuntimeError as ex:
                    assert 'number of non-zero array elements changed during function execution' in str(ex)

    run_threaded(func, max_workers=10, pass_count=True, outer_iterations=5)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\greenlet\tests\test_tracing.py ===
from __future__ import print_function
import sys
import greenlet
import unittest

from . import TestCase
from . import PY312

# https://discuss.python.org/t/cpython-3-12-greenlet-and-tracing-profiling-how-to-not-crash-and-get-correct-results/33144/2
DEBUG_BUILD_PY312 = (
    PY312 and hasattr(sys, 'gettotalrefcount'),
    "Broken on debug builds of Python 3.12"
)

class SomeError(Exception):
    pass

class GreenletTracer(object):
    oldtrace = None

    def __init__(self, error_on_trace=False):
        self.actions = []
        self.error_on_trace = error_on_trace

    def __call__(self, *args):
        self.actions.append(args)
        if self.error_on_trace:
            raise SomeError

    def __enter__(self):
        self.oldtrace = greenlet.settrace(self)
        return self.actions

    def __exit__(self, *args):
        greenlet.settrace(self.oldtrace)


class TestGreenletTracing(TestCase):
    """
    Tests of ``greenlet.settrace()``
    """

    def test_a_greenlet_tracing(self):
        main = greenlet.getcurrent()
        def dummy():
            pass
        def dummyexc():
            raise SomeError()

        with GreenletTracer() as actions:
            g1 = greenlet.greenlet(dummy)
            g1.switch()
            g2 = greenlet.greenlet(dummyexc)
            self.assertRaises(SomeError, g2.switch)

        self.assertEqual(actions, [
            ('switch', (main, g1)),
            ('switch', (g1, main)),
            ('switch', (main, g2)),
            ('throw', (g2, main)),
        ])

    def test_b_exception_disables_tracing(self):
        main = greenlet.getcurrent()
        def dummy():
            main.switch()
        g = greenlet.greenlet(dummy)
        g.switch()
        with GreenletTracer(error_on_trace=True) as actions:
            self.assertRaises(SomeError, g.switch)
            self.assertEqual(greenlet.gettrace(), None)

        self.assertEqual(actions, [
            ('switch', (main, g)),
        ])

    def test_set_same_tracer_twice(self):
        # https://github.com/python-greenlet/greenlet/issues/332
        # Our logic in asserting that the tracefunction should
        # gain a reference was incorrect if the same tracefunction was set
        # twice.
        tracer = GreenletTracer()
        with tracer:
            greenlet.settrace(tracer)


class PythonTracer(object):
    oldtrace = None

    def __init__(self):
        self.actions = []

    def __call__(self, frame, event, arg):
        # Record the co_name so we have an idea what function we're in.
        self.actions.append((event, frame.f_code.co_name))

    def __enter__(self):
        self.oldtrace = sys.setprofile(self)
        return self.actions

    def __exit__(self, *args):
        sys.setprofile(self.oldtrace)

def tpt_callback():
    return 42

class TestPythonTracing(TestCase):
    """
    Tests of the interaction of ``sys.settrace()``
    with greenlet facilities.

    NOTE: Most of this is probably CPython specific.
    """

    maxDiff = None

    def test_trace_events_trivial(self):
        with PythonTracer() as actions:
            tpt_callback()
        # If we use the sys.settrace instead of setprofile, we get
        # this:

        # self.assertEqual(actions, [
        #     ('call', 'tpt_callback'),
        #     ('call', '__exit__'),
        # ])

        self.assertEqual(actions, [
            ('return', '__enter__'),
            ('call', 'tpt_callback'),
            ('return', 'tpt_callback'),
            ('call', '__exit__'),
            ('c_call', '__exit__'),
        ])

    def _trace_switch(self, glet):
        with PythonTracer() as actions:
            glet.switch()
        return actions

    def _check_trace_events_func_already_set(self, glet):
        actions = self._trace_switch(glet)
        self.assertEqual(actions, [
            ('return', '__enter__'),
            ('c_call', '_trace_switch'),
            ('call', 'run'),
            ('call', 'tpt_callback'),
            ('return', 'tpt_callback'),
            ('return', 'run'),
            ('c_return', '_trace_switch'),
            ('call', '__exit__'),
            ('c_call', '__exit__'),
        ])

    def test_trace_events_into_greenlet_func_already_set(self):
        def run():
            return tpt_callback()

        self._check_trace_events_func_already_set(greenlet.greenlet(run))

    def test_trace_events_into_greenlet_subclass_already_set(self):
        class X(greenlet.greenlet):
            def run(self):
                return tpt_callback()
        self._check_trace_events_func_already_set(X())

    def _check_trace_events_from_greenlet_sets_profiler(self, g, tracer):
        g.switch()
        tpt_callback()
        tracer.__exit__()
        self.assertEqual(tracer.actions, [
            ('return', '__enter__'),
            ('call', 'tpt_callback'),
            ('return', 'tpt_callback'),
            ('return', 'run'),
            ('call', 'tpt_callback'),
            ('return', 'tpt_callback'),
            ('call', '__exit__'),
            ('c_call', '__exit__'),
        ])


    def test_trace_events_from_greenlet_func_sets_profiler(self):
        tracer = PythonTracer()
        def run():
            tracer.__enter__()
            return tpt_callback()

        self._check_trace_events_from_greenlet_sets_profiler(greenlet.greenlet(run),
                                                             tracer)

    def test_trace_events_from_greenlet_subclass_sets_profiler(self):
        tracer = PythonTracer()
        class X(greenlet.greenlet):
            def run(self):
                tracer.__enter__()
                return tpt_callback()

        self._check_trace_events_from_greenlet_sets_profiler(X(), tracer)

    @unittest.skipIf(*DEBUG_BUILD_PY312)
    def test_trace_events_multiple_greenlets_switching(self):
        tracer = PythonTracer()

        g1 = None
        g2 = None

        def g1_run():
            tracer.__enter__()
            tpt_callback()
            g2.switch()
            tpt_callback()
            return 42

        def g2_run():
            tpt_callback()
            tracer.__exit__()
            tpt_callback()
            g1.switch()

        g1 = greenlet.greenlet(g1_run)
        g2 = greenlet.greenlet(g2_run)

        x = g1.switch()
        self.assertEqual(x, 42)
        tpt_callback() # ensure not in the trace
        self.assertEqual(tracer.actions, [
            ('return', '__enter__'),
            ('call', 'tpt_callback'),
            ('return', 'tpt_callback'),
            ('c_call', 'g1_run'),
            ('call', 'g2_run'),
            ('call', 'tpt_callback'),
            ('return', 'tpt_callback'),
            ('call', '__exit__'),
            ('c_call', '__exit__'),
        ])

    @unittest.skipIf(*DEBUG_BUILD_PY312)
    def test_trace_events_multiple_greenlets_switching_siblings(self):
        # Like the first version, but get both greenlets running first
        # as "siblings" and then establish the tracing.
        tracer = PythonTracer()

        g1 = None
        g2 = None

        def g1_run():
            greenlet.getcurrent().parent.switch()
            tracer.__enter__()
            tpt_callback()
            g2.switch()
            tpt_callback()
            return 42

        def g2_run():
            greenlet.getcurrent().parent.switch()

            tpt_callback()
            tracer.__exit__()
            tpt_callback()
            g1.switch()

        g1 = greenlet.greenlet(g1_run)
        g2 = greenlet.greenlet(g2_run)

        # Start g1
        g1.switch()
        # And it immediately returns control to us.
        # Start g2
        g2.switch()
        # Which also returns. Now kick of the real part of the
        # test.
        x = g1.switch()
        self.assertEqual(x, 42)

        tpt_callback() # ensure not in the trace
        self.assertEqual(tracer.actions, [
            ('return', '__enter__'),
            ('call', 'tpt_callback'),
            ('return', 'tpt_callback'),
            ('c_call', 'g1_run'),
            ('call', 'tpt_callback'),
            ('return', 'tpt_callback'),
            ('call', '__exit__'),
            ('c_call', '__exit__'),
        ])


if __name__ == '__main__':
    unittest.main()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\mpmath\tests\test_hp.py ===
"""
Check that the output from irrational functions is accurate for
high-precision input, from 5 to 200 digits. The reference values were
verified with Mathematica.
"""

import time
from mpmath import *

precs = [5, 15, 28, 35, 57, 80, 100, 150, 200]

# sqrt(3) + pi/2
a = \
"3.302847134363773912758768033145623809041389953497933538543279275605"\
"841220051904536395163599428307109666700184672047856353516867399774243594"\
"67433521615861420725323528325327484262075464241255915238845599752675"

# e + 1/euler**2
b = \
"5.719681166601007617111261398629939965860873957353320734275716220045750"\
"31474116300529519620938123730851145473473708966080207482581266469342214"\
"824842256999042984813905047895479210702109260221361437411947323431"

# sqrt(a)
sqrt_a = \
"1.817373691447021556327498239690365674922395036495564333152483422755"\
"144321726165582817927383239308173567921345318453306994746434073691275094"\
"484777905906961689902608644112196725896908619756404253109722911487"

# sqrt(a+b*i).real
sqrt_abi_real = \
"2.225720098415113027729407777066107959851146508557282707197601407276"\
"89160998185797504198062911768240808839104987021515555650875977724230130"\
"3584116233925658621288393930286871862273400475179312570274423840384"

# sqrt(a+b*i).imag
sqrt_abi_imag = \
"1.2849057639084690902371581529110949983261182430040898147672052833653668"\
"0629534491275114877090834296831373498336559849050755848611854282001250"\
"1924311019152914021365263161630765255610885489295778894976075186"

# log(a)
log_a = \
"1.194784864491089550288313512105715261520511949410072046160598707069"\
"4336653155025770546309137440687056366757650909754708302115204338077595203"\
"83005773986664564927027147084436553262269459110211221152925732612"

# log(a+b*i).real
log_abi_real = \
"1.8877985921697018111624077550443297276844736840853590212962006811663"\
"04949387789489704203167470111267581371396245317618589339274243008242708"\
"014251531496104028712866224020066439049377679709216784954509456421"

# log(a+b*i).imag
log_abi_imag = \
"1.0471204952840802663567714297078763189256357109769672185219334169734948"\
"4265809854092437285294686651806426649541504240470168212723133326542181"\
"8300136462287639956713914482701017346851009323172531601894918640"

# exp(a)
exp_a = \
"27.18994224087168661137253262213293847994194869430518354305430976149"\
"382792035050358791398632888885200049857986258414049540376323785711941636"\
"100358982497583832083513086941635049329804685212200507288797531143"

# exp(a+b*i).real
exp_abi_real = \
"22.98606617170543596386921087657586890620262522816912505151109385026"\
"40160179326569526152851983847133513990281518417211964710397233157168852"\
"4963130831190142571659948419307628119985383887599493378056639916701"

# exp(a+b*i).imag
exp_abi_imag = \
"-14.523557450291489727214750571590272774669907424478129280902375851196283"\
"3377162379031724734050088565710975758824441845278120105728824497308303"\
"6065619788140201636218705414429933685889542661364184694108251449"

# a**b
pow_a_b = \
"928.7025342285568142947391505837660251004990092821305668257284426997"\
"361966028275685583421197860603126498884545336686124793155581311527995550"\
"580229264427202446131740932666832138634013168125809402143796691154"

# (a**(a+b*i)).real
pow_a_abi_real = \
"44.09156071394489511956058111704382592976814280267142206420038656267"\
"67707916510652790502399193109819563864568986234654864462095231138500505"\
"8197456514795059492120303477512711977915544927440682508821426093455"

# (a**(a+b*i)).imag
pow_a_abi_imag = \
"27.069371511573224750478105146737852141664955461266218367212527612279886"\
"9322304536553254659049205414427707675802193810711302947536332040474573"\
"8166261217563960235014674118610092944307893857862518964990092301"

# ((a+b*i)**(a+b*i)).real
pow_abi_abi_real = \
"-0.15171310677859590091001057734676423076527145052787388589334350524"\
"8084195882019497779202452975350579073716811284169068082670778986235179"\
"0813026562962084477640470612184016755250592698408112493759742219150452"\

# ((a+b*i)**(a+b*i)).imag
pow_abi_abi_imag = \
"1.2697592504953448936553147870155987153192995316950583150964099070426"\
"4736837932577176947632535475040521749162383347758827307504526525647759"\
"97547638617201824468382194146854367480471892602963428122896045019902"

# sin(a)
sin_a = \
"-0.16055653857469062740274792907968048154164433772938156243509084009"\
"38437090841460493108570147191289893388608611542655654723437248152535114"\
"528368009465836614227575701220612124204622383149391870684288862269631"

# sin(1000*a)
sin_1000a = \
"-0.85897040577443833776358106803777589664322997794126153477060795801"\
"09151695416961724733492511852267067419573754315098042850381158563024337"\
"216458577140500488715469780315833217177634490142748614625281171216863"

# sin(a+b*i)
sin_abi_real = \
"-24.4696999681556977743346798696005278716053366404081910969773939630"\
"7149215135459794473448465734589287491880563183624997435193637389884206"\
"02151395451271809790360963144464736839412254746645151672423256977064"

sin_abi_imag = \
"-150.42505378241784671801405965872972765595073690984080160750785565810981"\
"8314482499135443827055399655645954830931316357243750839088113122816583"\
"7169201254329464271121058839499197583056427233866320456505060735"

# cos
cos_a = \
"-0.98702664499035378399332439243967038895709261414476495730788864004"\
"05406821549361039745258003422386169330787395654908532996287293003581554"\
"257037193284199198069707141161341820684198547572456183525659969145501"

cos_1000a = \
"-0.51202523570982001856195696460663971099692261342827540426136215533"\
"52686662667660613179619804463250686852463876088694806607652218586060613"\
"951310588158830695735537073667299449753951774916401887657320950496820"

# tan
tan_a = \
"0.162666873675188117341401059858835168007137819495998960250142156848"\
"639654718809412181543343168174807985559916643549174530459883826451064966"\
"7996119428949951351938178809444268785629011625179962457123195557310"

tan_abi_real = \
"6.822696615947538488826586186310162599974827139564433912601918442911"\
"1026830824380070400102213741875804368044342309515353631134074491271890"\
"467615882710035471686578162073677173148647065131872116479947620E-6"

tan_abi_imag = \
"0.9999795833048243692245661011298447587046967777739649018690797625964167"\
"1446419978852235960862841608081413169601038230073129482874832053357571"\
"62702259309150715669026865777947502665936317953101462202542168429"


def test_hp():
    for dps in precs:
        mp.dps = dps + 8
        aa = mpf(a)
        bb = mpf(b)
        a1000 = 1000*mpf(a)
        abi = mpc(aa, bb)
        mp.dps = dps
        assert (sqrt(3) + pi/2).ae(aa)
        assert (e + 1/euler**2).ae(bb)

        assert sqrt(aa).ae(mpf(sqrt_a))
        assert sqrt(abi).ae(mpc(sqrt_abi_real, sqrt_abi_imag))

        assert log(aa).ae(mpf(log_a))
        assert log(abi).ae(mpc(log_abi_real, log_abi_imag))

        assert exp(aa).ae(mpf(exp_a))
        assert exp(abi).ae(mpc(exp_abi_real, exp_abi_imag))

        assert (aa**bb).ae(mpf(pow_a_b))
        assert (aa**abi).ae(mpc(pow_a_abi_real, pow_a_abi_imag))
        assert (abi**abi).ae(mpc(pow_abi_abi_real, pow_abi_abi_imag))

        assert sin(a).ae(mpf(sin_a))
        assert sin(a1000).ae(mpf(sin_1000a))
        assert sin(abi).ae(mpc(sin_abi_real, sin_abi_imag))

        assert cos(a).ae(mpf(cos_a))
        assert cos(a1000).ae(mpf(cos_1000a))

        assert tan(a).ae(mpf(tan_a))
        assert tan(abi).ae(mpc(tan_abi_real, tan_abi_imag))

        # check that complex cancellation is avoided so that both
        # real and imaginary parts have high relative accuracy.
        # abs_eps should be 0, but has to be set to 1e-205 to pass the
        # 200-digit case, probably due to slight inaccuracy in the
        # precomputed input
        assert (tan(abi).real).ae(mpf(tan_abi_real), abs_eps=1e-205)
        assert (tan(abi).imag).ae(mpf(tan_abi_imag), abs_eps=1e-205)
    mp.dps = 460
    assert str(log(3))[-20:] == '02166121184001409826'
    mp.dps = 15

# Since str(a) can differ in the last digit from rounded a, and I want
# to compare the last digits of big numbers with the results in Mathematica,
# I made this hack to get the last 20 digits of rounded a

def last_digits(a):
    r = repr(a)
    s = str(a)
    #dps = mp.dps
    #mp.dps += 3
    m = 10
    r = r.replace(s[:-m],'')
    r = r.replace("mpf('",'').replace("')",'')
    num0 = 0
    for c in r:
        if c == '0':
            num0 += 1
        else:
            break
    b = float(int(r))/10**(len(r) - m)
    if b >= 10**m - 0.5:  # pragma: no cover
        raise NotImplementedError
    n = int(round(b))
    sn = str(n)
    s = s[:-m] + '0'*num0 + sn
    return s[-20:]

# values checked with Mathematica
def test_log_hp():
    mp.dps = 2000
    a = mpf(10)**15000/3
    r = log(a)
    res = last_digits(r)
    # Mathematica N[Log[10^15000/3], 2000]
    # ...7443804441768333470331
    assert res == '43804441768333470331'

    # see issue 145
    r = log(mpf(3)/2)
    # Mathematica N[Log[3/2], 2000]
    # ...69653749808140753263288
    res = last_digits(r)
    assert res == '53749808140753263288'

    mp.dps = 10000
    r = log(2)
    res = last_digits(r)
    # Mathematica  N[Log[2], 10000]
    # ...695615913401856601359655561
    assert res == '13401856601359655561'
    r = log(mpf(10)**10/3)
    res = last_digits(r)
    # Mathematica N[Log[10^10/3], 10000]
    # ...587087654020631943060007154
    assert res == '54020631943060007154', res
    r = log(mpf(10)**100/3)
    res = last_digits(r)
    # Mathematica N[Log[10^100/3], 10000]
    # ,,,59246336539088351652334666
    assert res == '36539088351652334666', res
    mp.dps += 10
    a = 1 - mpf(1)/10**10
    mp.dps -= 10
    r = log(a)
    res = last_digits(r)
    # ...3310334360482956137216724048322957404
    # 372167240483229574038733026370
    # Mathematica N[Log[1 - 10^-10]*10^10, 10000]
    # ...60482956137216724048322957404
    assert res == '37216724048322957404', res
    mp.dps = 10000
    mp.dps += 100
    a = 1 + mpf(1)/10**100
    mp.dps -= 100

    r = log(a)
    res = last_digits(+r)
    # Mathematica N[Log[1 + 10^-100]*10^10, 10030]
    # ...3994733877377412241546890854692521568292338268273 10^-91
    assert res == '39947338773774122415', res

    mp.dps = 15

def test_exp_hp():
    mp.dps = 4000
    r = exp(mpf(1)/10)
    # IntegerPart[N[Exp[1/10] * 10^4000, 4000]]
    # ...92167105162069688129
    assert int(r * 10**mp.dps) % 10**20 == 92167105162069688129

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\indexes\timedeltas\test_constructors.py ===
from datetime import timedelta

import numpy as np
import pytest

import pandas as pd
from pandas import (
    Timedelta,
    TimedeltaIndex,
    timedelta_range,
    to_timedelta,
)
import pandas._testing as tm
from pandas.core.arrays.timedeltas import TimedeltaArray


class TestTimedeltaIndex:
    def test_closed_deprecated(self):
        # GH#52628
        msg = "The 'closed' keyword"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            TimedeltaIndex([], closed=True)

    def test_array_of_dt64_nat_raises(self):
        # GH#39462
        nat = np.datetime64("NaT", "ns")
        arr = np.array([nat], dtype=object)

        msg = "Invalid type for timedelta scalar"
        with pytest.raises(TypeError, match=msg):
            TimedeltaIndex(arr)

        with pytest.raises(TypeError, match=msg):
            TimedeltaArray._from_sequence(arr, dtype="m8[ns]")

        with pytest.raises(TypeError, match=msg):
            to_timedelta(arr)

    @pytest.mark.parametrize("unit", ["Y", "y", "M"])
    def test_unit_m_y_raises(self, unit):
        msg = "Units 'M', 'Y', and 'y' are no longer supported"
        depr_msg = "The 'unit' keyword in TimedeltaIndex construction is deprecated"
        with pytest.raises(ValueError, match=msg):
            with tm.assert_produces_warning(FutureWarning, match=depr_msg):
                TimedeltaIndex([1, 3, 7], unit)

    def test_int64_nocopy(self):
        # GH#23539 check that a copy isn't made when we pass int64 data
        #  and copy=False
        arr = np.arange(10, dtype=np.int64)
        tdi = TimedeltaIndex(arr, copy=False)
        assert tdi._data._ndarray.base is arr

    def test_infer_from_tdi(self):
        # GH#23539
        # fast-path for inferring a frequency if the passed data already
        #  has one
        tdi = timedelta_range("1 second", periods=10**7, freq="1s")

        result = TimedeltaIndex(tdi, freq="infer")
        assert result.freq == tdi.freq

        # check that inferred_freq was not called by checking that the
        #  value has not been cached
        assert "inferred_freq" not in getattr(result, "_cache", {})

    def test_infer_from_tdi_mismatch(self):
        # GH#23539
        # fast-path for invalidating a frequency if the passed data already
        #  has one and it does not match the `freq` input
        tdi = timedelta_range("1 second", periods=100, freq="1s")

        depr_msg = "TimedeltaArray.__init__ is deprecated"
        msg = (
            "Inferred frequency .* from passed values does "
            "not conform to passed frequency"
        )
        with pytest.raises(ValueError, match=msg):
            TimedeltaIndex(tdi, freq="D")

        with pytest.raises(ValueError, match=msg):
            # GH#23789
            with tm.assert_produces_warning(FutureWarning, match=depr_msg):
                TimedeltaArray(tdi, freq="D")

        with pytest.raises(ValueError, match=msg):
            TimedeltaIndex(tdi._data, freq="D")

        with pytest.raises(ValueError, match=msg):
            with tm.assert_produces_warning(FutureWarning, match=depr_msg):
                TimedeltaArray(tdi._data, freq="D")

    def test_dt64_data_invalid(self):
        # GH#23539
        # passing tz-aware DatetimeIndex raises, naive or ndarray[datetime64]
        #  raise as of GH#29794
        dti = pd.date_range("2016-01-01", periods=3)

        msg = "cannot be converted to timedelta64"
        with pytest.raises(TypeError, match=msg):
            TimedeltaIndex(dti.tz_localize("Europe/Brussels"))

        with pytest.raises(TypeError, match=msg):
            TimedeltaIndex(dti)

        with pytest.raises(TypeError, match=msg):
            TimedeltaIndex(np.asarray(dti))

    def test_float64_ns_rounded(self):
        # GH#23539 without specifying a unit, floats are regarded as nanos,
        #  and fractional portions are truncated
        tdi = TimedeltaIndex([2.3, 9.7])
        expected = TimedeltaIndex([2, 9])
        tm.assert_index_equal(tdi, expected)

        # integral floats are non-lossy
        tdi = TimedeltaIndex([2.0, 9.0])
        expected = TimedeltaIndex([2, 9])
        tm.assert_index_equal(tdi, expected)

        # NaNs get converted to NaT
        tdi = TimedeltaIndex([2.0, np.nan])
        expected = TimedeltaIndex([Timedelta(nanoseconds=2), pd.NaT])
        tm.assert_index_equal(tdi, expected)

    def test_float64_unit_conversion(self):
        # GH#23539
        tdi = to_timedelta([1.5, 2.25], unit="D")
        expected = TimedeltaIndex([Timedelta(days=1.5), Timedelta(days=2.25)])
        tm.assert_index_equal(tdi, expected)

    def test_construction_base_constructor(self):
        arr = [Timedelta("1 days"), pd.NaT, Timedelta("3 days")]
        tm.assert_index_equal(pd.Index(arr), TimedeltaIndex(arr))
        tm.assert_index_equal(pd.Index(np.array(arr)), TimedeltaIndex(np.array(arr)))

        arr = [np.nan, pd.NaT, Timedelta("1 days")]
        tm.assert_index_equal(pd.Index(arr), TimedeltaIndex(arr))
        tm.assert_index_equal(pd.Index(np.array(arr)), TimedeltaIndex(np.array(arr)))

    @pytest.mark.filterwarnings(
        "ignore:The 'unit' keyword in TimedeltaIndex construction:FutureWarning"
    )
    def test_constructor(self):
        expected = TimedeltaIndex(
            [
                "1 days",
                "1 days 00:00:05",
                "2 days",
                "2 days 00:00:02",
                "0 days 00:00:03",
            ]
        )
        result = TimedeltaIndex(
            [
                "1 days",
                "1 days, 00:00:05",
                np.timedelta64(2, "D"),
                timedelta(days=2, seconds=2),
                pd.offsets.Second(3),
            ]
        )
        tm.assert_index_equal(result, expected)

        expected = TimedeltaIndex(
            ["0 days 00:00:00", "0 days 00:00:01", "0 days 00:00:02"]
        )
        result = TimedeltaIndex(range(3), unit="s")
        tm.assert_index_equal(result, expected)
        expected = TimedeltaIndex(
            ["0 days 00:00:00", "0 days 00:00:05", "0 days 00:00:09"]
        )
        result = TimedeltaIndex([0, 5, 9], unit="s")
        tm.assert_index_equal(result, expected)
        expected = TimedeltaIndex(
            ["0 days 00:00:00.400", "0 days 00:00:00.450", "0 days 00:00:01.200"]
        )
        result = TimedeltaIndex([400, 450, 1200], unit="ms")
        tm.assert_index_equal(result, expected)

    def test_constructor_iso(self):
        # GH #21877
        expected = timedelta_range("1s", periods=9, freq="s")
        durations = [f"P0DT0H0M{i}S" for i in range(1, 10)]
        result = to_timedelta(durations)
        tm.assert_index_equal(result, expected)

    def test_timedelta_range_fractional_period(self):
        msg = "Non-integer 'periods' in pd.date_range, pd.timedelta_range"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            rng = timedelta_range("1 days", periods=10.5)
        exp = timedelta_range("1 days", periods=10)
        tm.assert_index_equal(rng, exp)

    def test_constructor_coverage(self):
        msg = "periods must be a number, got foo"
        with pytest.raises(TypeError, match=msg):
            timedelta_range(start="1 days", periods="foo", freq="D")

        msg = (
            r"TimedeltaIndex\(\.\.\.\) must be called with a collection of some kind, "
            "'1 days' was passed"
        )
        with pytest.raises(TypeError, match=msg):
            TimedeltaIndex("1 days")

        # generator expression
        gen = (timedelta(i) for i in range(10))
        result = TimedeltaIndex(gen)
        expected = TimedeltaIndex([timedelta(i) for i in range(10)])
        tm.assert_index_equal(result, expected)

        # NumPy string array
        strings = np.array(["1 days", "2 days", "3 days"])
        result = TimedeltaIndex(strings)
        expected = to_timedelta([1, 2, 3], unit="d")
        tm.assert_index_equal(result, expected)

        from_ints = TimedeltaIndex(expected.asi8)
        tm.assert_index_equal(from_ints, expected)

        # non-conforming freq
        msg = (
            "Inferred frequency None from passed values does not conform to "
            "passed frequency D"
        )
        with pytest.raises(ValueError, match=msg):
            TimedeltaIndex(["1 days", "2 days", "4 days"], freq="D")

        msg = (
            "Of the four parameters: start, end, periods, and freq, exactly "
            "three must be specified"
        )
        with pytest.raises(ValueError, match=msg):
            timedelta_range(periods=10, freq="D")

    def test_constructor_name(self):
        idx = timedelta_range(start="1 days", periods=1, freq="D", name="TEST")
        assert idx.name == "TEST"

        # GH10025
        idx2 = TimedeltaIndex(idx, name="something else")
        assert idx2.name == "something else"

    def test_constructor_no_precision_raises(self):
        # GH-24753, GH-24739

        msg = "with no precision is not allowed"
        with pytest.raises(ValueError, match=msg):
            TimedeltaIndex(["2000"], dtype="timedelta64")

        msg = "The 'timedelta64' dtype has no unit. Please pass in"
        with pytest.raises(ValueError, match=msg):
            pd.Index(["2000"], dtype="timedelta64")

    def test_constructor_wrong_precision_raises(self):
        msg = "Supported timedelta64 resolutions are 's', 'ms', 'us', 'ns'"
        with pytest.raises(ValueError, match=msg):
            TimedeltaIndex(["2000"], dtype="timedelta64[D]")

        # "timedelta64[us]" was unsupported pre-2.0, but now this works.
        tdi = TimedeltaIndex(["2000"], dtype="timedelta64[us]")
        assert tdi.dtype == "m8[us]"

    def test_explicit_none_freq(self):
        # Explicitly passing freq=None is respected
        tdi = timedelta_range(1, periods=5)
        assert tdi.freq is not None

        result = TimedeltaIndex(tdi, freq=None)
        assert result.freq is None

        result = TimedeltaIndex(tdi._data, freq=None)
        assert result.freq is None

        msg = "TimedeltaArray.__init__ is deprecated"
        with tm.assert_produces_warning(FutureWarning, match=msg):
            tda = TimedeltaArray(tdi, freq=None)
        assert tda.freq is None

    def test_from_categorical(self):
        tdi = timedelta_range(1, periods=5)

        cat = pd.Categorical(tdi)

        result = TimedeltaIndex(cat)
        tm.assert_index_equal(result, tdi)

        ci = pd.CategoricalIndex(tdi)
        result = TimedeltaIndex(ci)
        tm.assert_index_equal(result, tdi)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\extension\base\ops.py ===
from __future__ import annotations

from typing import final

import numpy as np
import pytest

from pandas.core.dtypes.common import is_string_dtype

import pandas as pd
import pandas._testing as tm
from pandas.core import ops


class BaseOpsUtil:
    series_scalar_exc: type[Exception] | None = TypeError
    frame_scalar_exc: type[Exception] | None = TypeError
    series_array_exc: type[Exception] | None = TypeError
    divmod_exc: type[Exception] | None = TypeError

    def _get_expected_exception(
        self, op_name: str, obj, other
    ) -> type[Exception] | tuple[type[Exception], ...] | None:
        # Find the Exception, if any we expect to raise calling
        #  obj.__op_name__(other)

        # The self.obj_bar_exc pattern isn't great in part because it can depend
        #  on op_name or dtypes, but we use it here for backward-compatibility.
        if op_name in ["__divmod__", "__rdivmod__"]:
            result = self.divmod_exc
        elif isinstance(obj, pd.Series) and isinstance(other, pd.Series):
            result = self.series_array_exc
        elif isinstance(obj, pd.Series):
            result = self.series_scalar_exc
        else:
            result = self.frame_scalar_exc

        return result

    def _cast_pointwise_result(self, op_name: str, obj, other, pointwise_result):
        # In _check_op we check that the result of a pointwise operation
        #  (found via _combine) matches the result of the vectorized
        #  operation obj.__op_name__(other).
        #  In some cases pandas dtype inference on the scalar result may not
        #  give a matching dtype even if both operations are behaving "correctly".
        #  In these cases, do extra required casting here.
        return pointwise_result

    def get_op_from_name(self, op_name: str):
        return tm.get_op_from_name(op_name)

    # Subclasses are not expected to need to override check_opname, _check_op,
    #  _check_divmod_op, or _combine.
    #  Ideally any relevant overriding can be done in _cast_pointwise_result,
    #  get_op_from_name, and the specification of `exc`. If you find a use
    #  case that still requires overriding _check_op or _combine, please let
    #  us know at github.com/pandas-dev/pandas/issues
    @final
    def check_opname(self, ser: pd.Series, op_name: str, other):
        exc = self._get_expected_exception(op_name, ser, other)
        op = self.get_op_from_name(op_name)

        self._check_op(ser, op, other, op_name, exc)

    # see comment on check_opname
    @final
    def _combine(self, obj, other, op):
        if isinstance(obj, pd.DataFrame):
            if len(obj.columns) != 1:
                raise NotImplementedError
            expected = obj.iloc[:, 0].combine(other, op).to_frame()
        else:
            expected = obj.combine(other, op)
        return expected

    # see comment on check_opname
    @final
    def _check_op(
        self, ser: pd.Series, op, other, op_name: str, exc=NotImplementedError
    ):
        # Check that the Series/DataFrame arithmetic/comparison method matches
        #  the pointwise result from _combine.

        if exc is None:
            result = op(ser, other)
            expected = self._combine(ser, other, op)
            expected = self._cast_pointwise_result(op_name, ser, other, expected)
            assert isinstance(result, type(ser))
            tm.assert_equal(result, expected)
        else:
            with pytest.raises(exc):
                op(ser, other)

    # see comment on check_opname
    @final
    def _check_divmod_op(self, ser: pd.Series, op, other):
        # check that divmod behavior matches behavior of floordiv+mod
        if op is divmod:
            exc = self._get_expected_exception("__divmod__", ser, other)
        else:
            exc = self._get_expected_exception("__rdivmod__", ser, other)
        if exc is None:
            result_div, result_mod = op(ser, other)
            if op is divmod:
                expected_div, expected_mod = ser // other, ser % other
            else:
                expected_div, expected_mod = other // ser, other % ser
            tm.assert_series_equal(result_div, expected_div)
            tm.assert_series_equal(result_mod, expected_mod)
        else:
            with pytest.raises(exc):
                divmod(ser, other)


class BaseArithmeticOpsTests(BaseOpsUtil):
    """
    Various Series and DataFrame arithmetic ops methods.

    Subclasses supporting various ops should set the class variables
    to indicate that they support ops of that kind

    * series_scalar_exc = TypeError
    * frame_scalar_exc = TypeError
    * series_array_exc = TypeError
    * divmod_exc = TypeError
    """

    series_scalar_exc: type[Exception] | None = TypeError
    frame_scalar_exc: type[Exception] | None = TypeError
    series_array_exc: type[Exception] | None = TypeError
    divmod_exc: type[Exception] | None = TypeError

    def test_arith_series_with_scalar(self, data, all_arithmetic_operators):
        # series & scalar
        if all_arithmetic_operators == "__rmod__" and is_string_dtype(data.dtype):
            pytest.skip("Skip testing Python string formatting")

        op_name = all_arithmetic_operators
        ser = pd.Series(data)
        self.check_opname(ser, op_name, ser.iloc[0])

    def test_arith_frame_with_scalar(self, data, all_arithmetic_operators):
        # frame & scalar
        if all_arithmetic_operators == "__rmod__" and is_string_dtype(data.dtype):
            pytest.skip("Skip testing Python string formatting")

        op_name = all_arithmetic_operators
        df = pd.DataFrame({"A": data})
        self.check_opname(df, op_name, data[0])

    def test_arith_series_with_array(self, data, all_arithmetic_operators):
        # ndarray & other series
        op_name = all_arithmetic_operators
        ser = pd.Series(data)
        self.check_opname(ser, op_name, pd.Series([ser.iloc[0]] * len(ser)))

    def test_divmod(self, data):
        ser = pd.Series(data)
        self._check_divmod_op(ser, divmod, 1)
        self._check_divmod_op(1, ops.rdivmod, ser)

    def test_divmod_series_array(self, data, data_for_twos):
        ser = pd.Series(data)
        self._check_divmod_op(ser, divmod, data)

        other = data_for_twos
        self._check_divmod_op(other, ops.rdivmod, ser)

        other = pd.Series(other)
        self._check_divmod_op(other, ops.rdivmod, ser)

    def test_add_series_with_extension_array(self, data):
        # Check adding an ExtensionArray to a Series of the same dtype matches
        # the behavior of adding the arrays directly and then wrapping in a
        # Series.

        ser = pd.Series(data)

        exc = self._get_expected_exception("__add__", ser, data)
        if exc is not None:
            with pytest.raises(exc):
                ser + data
            return

        result = ser + data
        expected = pd.Series(data + data)
        tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize("box", [pd.Series, pd.DataFrame, pd.Index])
    @pytest.mark.parametrize(
        "op_name",
        [
            x
            for x in tm.arithmetic_dunder_methods + tm.comparison_dunder_methods
            if not x.startswith("__r")
        ],
    )
    def test_direct_arith_with_ndframe_returns_not_implemented(
        self, data, box, op_name
    ):
        # EAs should return NotImplemented for ops with Series/DataFrame/Index
        # Pandas takes care of unboxing the series and calling the EA's op.
        other = box(data)

        if hasattr(data, op_name):
            result = getattr(data, op_name)(other)
            assert result is NotImplemented


class BaseComparisonOpsTests(BaseOpsUtil):
    """Various Series and DataFrame comparison ops methods."""

    def _compare_other(self, ser: pd.Series, data, op, other):
        if op.__name__ in ["eq", "ne"]:
            # comparison should match point-wise comparisons
            result = op(ser, other)
            expected = ser.combine(other, op)
            expected = self._cast_pointwise_result(op.__name__, ser, other, expected)
            tm.assert_series_equal(result, expected)

        else:
            exc = None
            try:
                result = op(ser, other)
            except Exception as err:
                exc = err

            if exc is None:
                # Didn't error, then should match pointwise behavior
                expected = ser.combine(other, op)
                expected = self._cast_pointwise_result(
                    op.__name__, ser, other, expected
                )
                tm.assert_series_equal(result, expected)
            else:
                with pytest.raises(type(exc)):
                    ser.combine(other, op)

    def test_compare_scalar(self, data, comparison_op):
        ser = pd.Series(data)
        self._compare_other(ser, data, comparison_op, 0)

    def test_compare_array(self, data, comparison_op):
        ser = pd.Series(data)
        other = pd.Series([data[0]] * len(data), dtype=data.dtype)
        self._compare_other(ser, data, comparison_op, other)


class BaseUnaryOpsTests(BaseOpsUtil):
    def test_invert(self, data):
        ser = pd.Series(data, name="name")
        try:
            # 10 is an arbitrary choice here, just avoid iterating over
            #  the whole array to trim test runtime
            [~x for x in data[:10]]
        except TypeError:
            # scalars don't support invert -> we don't expect the vectorized
            #  operation to succeed
            with pytest.raises(TypeError):
                ~ser
            with pytest.raises(TypeError):
                ~data
        else:
            # Note we do not reuse the pointwise result to construct expected
            #  because python semantics for negating bools are weird see GH#54569
            result = ~ser
            expected = pd.Series(~data, name="name")
            tm.assert_series_equal(result, expected)

    @pytest.mark.parametrize("ufunc", [np.positive, np.negative, np.abs])
    def test_unary_ufunc_dunder_equivalence(self, data, ufunc):
        # the dunder __pos__ works if and only if np.positive works,
        #  same for __neg__/np.negative and __abs__/np.abs
        attr = {np.positive: "__pos__", np.negative: "__neg__", np.abs: "__abs__"}[
            ufunc
        ]

        exc = None
        try:
            result = getattr(data, attr)()
        except Exception as err:
            exc = err

            # if __pos__ raised, then so should the ufunc
            with pytest.raises((type(exc), TypeError)):
                ufunc(data)
        else:
            alt = ufunc(data)
            tm.assert_extension_array_equal(result, alt)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\indexes\multi\test_integrity.py ===
import re

import numpy as np
import pytest

from pandas._libs import index as libindex

from pandas.core.dtypes.cast import construct_1d_object_array_from_listlike

import pandas as pd
from pandas import (
    Index,
    IntervalIndex,
    MultiIndex,
    RangeIndex,
)
import pandas._testing as tm


def test_labels_dtypes():
    # GH 8456
    i = MultiIndex.from_tuples([("A", 1), ("A", 2)])
    assert i.codes[0].dtype == "int8"
    assert i.codes[1].dtype == "int8"

    i = MultiIndex.from_product([["a"], range(40)])
    assert i.codes[1].dtype == "int8"
    i = MultiIndex.from_product([["a"], range(400)])
    assert i.codes[1].dtype == "int16"
    i = MultiIndex.from_product([["a"], range(40000)])
    assert i.codes[1].dtype == "int32"

    i = MultiIndex.from_product([["a"], range(1000)])
    assert (i.codes[0] >= 0).all()
    assert (i.codes[1] >= 0).all()


def test_values_boxed():
    tuples = [
        (1, pd.Timestamp("2000-01-01")),
        (2, pd.NaT),
        (3, pd.Timestamp("2000-01-03")),
        (1, pd.Timestamp("2000-01-04")),
        (2, pd.Timestamp("2000-01-02")),
        (3, pd.Timestamp("2000-01-03")),
    ]
    result = MultiIndex.from_tuples(tuples)
    expected = construct_1d_object_array_from_listlike(tuples)
    tm.assert_numpy_array_equal(result.values, expected)
    # Check that code branches for boxed values produce identical results
    tm.assert_numpy_array_equal(result.values[:4], result[:4].values)


def test_values_multiindex_datetimeindex():
    # Test to ensure we hit the boxing / nobox part of MI.values
    ints = np.arange(10**18, 10**18 + 5)
    naive = pd.DatetimeIndex(ints)

    aware = pd.DatetimeIndex(ints, tz="US/Central")

    idx = MultiIndex.from_arrays([naive, aware])
    result = idx.values

    outer = pd.DatetimeIndex([x[0] for x in result])
    tm.assert_index_equal(outer, naive)

    inner = pd.DatetimeIndex([x[1] for x in result])
    tm.assert_index_equal(inner, aware)

    # n_lev > n_lab
    result = idx[:2].values

    outer = pd.DatetimeIndex([x[0] for x in result])
    tm.assert_index_equal(outer, naive[:2])

    inner = pd.DatetimeIndex([x[1] for x in result])
    tm.assert_index_equal(inner, aware[:2])


def test_values_multiindex_periodindex():
    # Test to ensure we hit the boxing / nobox part of MI.values
    ints = np.arange(2007, 2012)
    pidx = pd.PeriodIndex(ints, freq="D")

    idx = MultiIndex.from_arrays([ints, pidx])
    result = idx.values

    outer = Index([x[0] for x in result])
    tm.assert_index_equal(outer, Index(ints, dtype=np.int64))

    inner = pd.PeriodIndex([x[1] for x in result])
    tm.assert_index_equal(inner, pidx)

    # n_lev > n_lab
    result = idx[:2].values

    outer = Index([x[0] for x in result])
    tm.assert_index_equal(outer, Index(ints[:2], dtype=np.int64))

    inner = pd.PeriodIndex([x[1] for x in result])
    tm.assert_index_equal(inner, pidx[:2])


def test_consistency():
    # need to construct an overflow
    major_axis = list(range(70000))
    minor_axis = list(range(10))

    major_codes = np.arange(70000)
    minor_codes = np.repeat(range(10), 7000)

    # the fact that is works means it's consistent
    index = MultiIndex(
        levels=[major_axis, minor_axis], codes=[major_codes, minor_codes]
    )

    # inconsistent
    major_codes = np.array([0, 0, 1, 1, 1, 2, 2, 3, 3])
    minor_codes = np.array([0, 1, 0, 1, 1, 0, 1, 0, 1])
    index = MultiIndex(
        levels=[major_axis, minor_axis], codes=[major_codes, minor_codes]
    )

    assert index.is_unique is False


@pytest.mark.slow
def test_hash_collisions(monkeypatch):
    # non-smoke test that we don't get hash collisions
    size_cutoff = 50
    with monkeypatch.context() as m:
        m.setattr(libindex, "_SIZE_CUTOFF", size_cutoff)
        index = MultiIndex.from_product(
            [np.arange(8), np.arange(8)], names=["one", "two"]
        )
        result = index.get_indexer(index.values)
        tm.assert_numpy_array_equal(result, np.arange(len(index), dtype="intp"))

        for i in [0, 1, len(index) - 2, len(index) - 1]:
            result = index.get_loc(index[i])
            assert result == i


def test_dims():
    pass


def test_take_invalid_kwargs():
    vals = [["A", "B"], [pd.Timestamp("2011-01-01"), pd.Timestamp("2011-01-02")]]
    idx = MultiIndex.from_product(vals, names=["str", "dt"])
    indices = [1, 2]

    msg = r"take\(\) got an unexpected keyword argument 'foo'"
    with pytest.raises(TypeError, match=msg):
        idx.take(indices, foo=2)

    msg = "the 'out' parameter is not supported"
    with pytest.raises(ValueError, match=msg):
        idx.take(indices, out=indices)

    msg = "the 'mode' parameter is not supported"
    with pytest.raises(ValueError, match=msg):
        idx.take(indices, mode="clip")


def test_isna_behavior(idx):
    # should not segfault GH5123
    # NOTE: if MI representation changes, may make sense to allow
    # isna(MI)
    msg = "isna is not defined for MultiIndex"
    with pytest.raises(NotImplementedError, match=msg):
        pd.isna(idx)


def test_large_multiindex_error(monkeypatch):
    # GH12527
    size_cutoff = 50
    with monkeypatch.context() as m:
        m.setattr(libindex, "_SIZE_CUTOFF", size_cutoff)
        df_below_cutoff = pd.DataFrame(
            1,
            index=MultiIndex.from_product([[1, 2], range(size_cutoff - 1)]),
            columns=["dest"],
        )
        with pytest.raises(KeyError, match=r"^\(-1, 0\)$"):
            df_below_cutoff.loc[(-1, 0), "dest"]
        with pytest.raises(KeyError, match=r"^\(3, 0\)$"):
            df_below_cutoff.loc[(3, 0), "dest"]
        df_above_cutoff = pd.DataFrame(
            1,
            index=MultiIndex.from_product([[1, 2], range(size_cutoff + 1)]),
            columns=["dest"],
        )
        with pytest.raises(KeyError, match=r"^\(-1, 0\)$"):
            df_above_cutoff.loc[(-1, 0), "dest"]
        with pytest.raises(KeyError, match=r"^\(3, 0\)$"):
            df_above_cutoff.loc[(3, 0), "dest"]


def test_mi_hashtable_populated_attribute_error(monkeypatch):
    # GH 18165
    monkeypatch.setattr(libindex, "_SIZE_CUTOFF", 50)
    r = range(50)
    df = pd.DataFrame({"a": r, "b": r}, index=MultiIndex.from_arrays([r, r]))

    msg = "'Series' object has no attribute 'foo'"
    with pytest.raises(AttributeError, match=msg):
        df["a"].foo()


def test_can_hold_identifiers(idx):
    key = idx[0]
    assert idx._can_hold_identifiers_and_holds_name(key) is True


def test_metadata_immutable(idx):
    levels, codes = idx.levels, idx.codes
    # shouldn't be able to set at either the top level or base level
    mutable_regex = re.compile("does not support mutable operations")
    with pytest.raises(TypeError, match=mutable_regex):
        levels[0] = levels[0]
    with pytest.raises(TypeError, match=mutable_regex):
        levels[0][0] = levels[0][0]
    # ditto for labels
    with pytest.raises(TypeError, match=mutable_regex):
        codes[0] = codes[0]
    with pytest.raises(ValueError, match="assignment destination is read-only"):
        codes[0][0] = codes[0][0]
    # and for names
    names = idx.names
    with pytest.raises(TypeError, match=mutable_regex):
        names[0] = names[0]


def test_level_setting_resets_attributes():
    ind = MultiIndex.from_arrays([["A", "A", "B", "B", "B"], [1, 2, 1, 2, 3]])
    assert ind.is_monotonic_increasing
    ind = ind.set_levels([["A", "B"], [1, 3, 2]])
    # if this fails, probably didn't reset the cache correctly.
    assert not ind.is_monotonic_increasing


def test_rangeindex_fallback_coercion_bug():
    # GH 12893
    df1 = pd.DataFrame(np.arange(100).reshape((10, 10)))
    df2 = pd.DataFrame(np.arange(100).reshape((10, 10)))
    df = pd.concat(
        {"df1": df1.stack(future_stack=True), "df2": df2.stack(future_stack=True)},
        axis=1,
    )
    df.index.names = ["fizz", "buzz"]

    expected = pd.DataFrame(
        {"df2": np.arange(100), "df1": np.arange(100)},
        index=MultiIndex.from_product([range(10), range(10)], names=["fizz", "buzz"]),
    )
    tm.assert_frame_equal(df, expected, check_like=True)

    result = df.index.get_level_values("fizz")
    expected = Index(np.arange(10, dtype=np.int64), name="fizz").repeat(10)
    tm.assert_index_equal(result, expected)

    result = df.index.get_level_values("buzz")
    expected = Index(np.tile(np.arange(10, dtype=np.int64), 10), name="buzz")
    tm.assert_index_equal(result, expected)


def test_memory_usage(idx):
    result = idx.memory_usage()
    if len(idx):
        idx.get_loc(idx[0])
        result2 = idx.memory_usage()
        result3 = idx.memory_usage(deep=True)

        # RangeIndex, IntervalIndex
        # don't have engines
        if not isinstance(idx, (RangeIndex, IntervalIndex)):
            assert result2 > result

        if idx.inferred_type == "object":
            assert result3 > result2

    else:
        # we report 0 for no-length
        assert result == 0


def test_nlevels(idx):
    assert idx.nlevels == 2

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\io\formats\test_css.py ===
import pytest

from pandas.errors import CSSWarning

import pandas._testing as tm

from pandas.io.formats.css import CSSResolver


def assert_resolves(css, props, inherited=None):
    resolve = CSSResolver()
    actual = resolve(css, inherited=inherited)
    assert props == actual


def assert_same_resolution(css1, css2, inherited=None):
    resolve = CSSResolver()
    resolved1 = resolve(css1, inherited=inherited)
    resolved2 = resolve(css2, inherited=inherited)
    assert resolved1 == resolved2


@pytest.mark.parametrize(
    "name,norm,abnorm",
    [
        (
            "whitespace",
            "hello: world; foo: bar",
            " \t hello \t :\n  world \n  ;  \n foo: \tbar\n\n",
        ),
        ("case", "hello: world; foo: bar", "Hello: WORLD; foO: bar"),
        ("empty-decl", "hello: world; foo: bar", "; hello: world;; foo: bar;\n; ;"),
        ("empty-list", "", ";"),
    ],
)
def test_css_parse_normalisation(name, norm, abnorm):
    assert_same_resolution(norm, abnorm)


@pytest.mark.parametrize(
    "invalid_css,remainder",
    [
        # No colon
        ("hello-world", ""),
        ("border-style: solid; hello-world", "border-style: solid"),
        (
            "border-style: solid; hello-world; font-weight: bold",
            "border-style: solid; font-weight: bold",
        ),
        # Unclosed string fail
        # Invalid size
        ("font-size: blah", "font-size: 1em"),
        ("font-size: 1a2b", "font-size: 1em"),
        ("font-size: 1e5pt", "font-size: 1em"),
        ("font-size: 1+6pt", "font-size: 1em"),
        ("font-size: 1unknownunit", "font-size: 1em"),
        ("font-size: 10", "font-size: 1em"),
        ("font-size: 10 pt", "font-size: 1em"),
        # Too many args
        ("border-top: 1pt solid red green", "border-top: 1pt solid green"),
    ],
)
def test_css_parse_invalid(invalid_css, remainder):
    with tm.assert_produces_warning(CSSWarning):
        assert_same_resolution(invalid_css, remainder)


@pytest.mark.parametrize(
    "shorthand,expansions",
    [
        ("margin", ["margin-top", "margin-right", "margin-bottom", "margin-left"]),
        ("padding", ["padding-top", "padding-right", "padding-bottom", "padding-left"]),
        (
            "border-width",
            [
                "border-top-width",
                "border-right-width",
                "border-bottom-width",
                "border-left-width",
            ],
        ),
        (
            "border-color",
            [
                "border-top-color",
                "border-right-color",
                "border-bottom-color",
                "border-left-color",
            ],
        ),
        (
            "border-style",
            [
                "border-top-style",
                "border-right-style",
                "border-bottom-style",
                "border-left-style",
            ],
        ),
    ],
)
def test_css_side_shorthands(shorthand, expansions):
    top, right, bottom, left = expansions

    assert_resolves(
        f"{shorthand}: 1pt", {top: "1pt", right: "1pt", bottom: "1pt", left: "1pt"}
    )

    assert_resolves(
        f"{shorthand}: 1pt 4pt", {top: "1pt", right: "4pt", bottom: "1pt", left: "4pt"}
    )

    assert_resolves(
        f"{shorthand}: 1pt 4pt 2pt",
        {top: "1pt", right: "4pt", bottom: "2pt", left: "4pt"},
    )

    assert_resolves(
        f"{shorthand}: 1pt 4pt 2pt 0pt",
        {top: "1pt", right: "4pt", bottom: "2pt", left: "0pt"},
    )

    with tm.assert_produces_warning(CSSWarning):
        assert_resolves(f"{shorthand}: 1pt 1pt 1pt 1pt 1pt", {})


@pytest.mark.parametrize(
    "shorthand,sides",
    [
        ("border-top", ["top"]),
        ("border-right", ["right"]),
        ("border-bottom", ["bottom"]),
        ("border-left", ["left"]),
        ("border", ["top", "right", "bottom", "left"]),
    ],
)
def test_css_border_shorthand_sides(shorthand, sides):
    def create_border_dict(sides, color=None, style=None, width=None):
        resolved = {}
        for side in sides:
            if color:
                resolved[f"border-{side}-color"] = color
            if style:
                resolved[f"border-{side}-style"] = style
            if width:
                resolved[f"border-{side}-width"] = width
        return resolved

    assert_resolves(
        f"{shorthand}: 1pt red solid", create_border_dict(sides, "red", "solid", "1pt")
    )


@pytest.mark.parametrize(
    "prop, expected",
    [
        ("1pt red solid", ("red", "solid", "1pt")),
        ("red 1pt solid", ("red", "solid", "1pt")),
        ("red solid 1pt", ("red", "solid", "1pt")),
        ("solid 1pt red", ("red", "solid", "1pt")),
        ("red solid", ("red", "solid", "1.500000pt")),
        # Note: color=black is not CSS conforming
        # (See https://drafts.csswg.org/css-backgrounds/#border-shorthands)
        ("1pt solid", ("black", "solid", "1pt")),
        ("1pt red", ("red", "none", "1pt")),
        ("red", ("red", "none", "1.500000pt")),
        ("1pt", ("black", "none", "1pt")),
        ("solid", ("black", "solid", "1.500000pt")),
        # Sizes
        ("1em", ("black", "none", "12pt")),
    ],
)
def test_css_border_shorthands(prop, expected):
    color, style, width = expected

    assert_resolves(
        f"border-left: {prop}",
        {
            "border-left-color": color,
            "border-left-style": style,
            "border-left-width": width,
        },
    )


@pytest.mark.parametrize(
    "style,inherited,equiv",
    [
        ("margin: 1px; margin: 2px", "", "margin: 2px"),
        ("margin: 1px", "margin: 2px", "margin: 1px"),
        ("margin: 1px; margin: inherit", "margin: 2px", "margin: 2px"),
        (
            "margin: 1px; margin-top: 2px",
            "",
            "margin-left: 1px; margin-right: 1px; "
            "margin-bottom: 1px; margin-top: 2px",
        ),
        ("margin-top: 2px", "margin: 1px", "margin: 1px; margin-top: 2px"),
        ("margin: 1px", "margin-top: 2px", "margin: 1px"),
        (
            "margin: 1px; margin-top: inherit",
            "margin: 2px",
            "margin: 1px; margin-top: 2px",
        ),
    ],
)
def test_css_precedence(style, inherited, equiv):
    resolve = CSSResolver()
    inherited_props = resolve(inherited)
    style_props = resolve(style, inherited=inherited_props)
    equiv_props = resolve(equiv)
    assert style_props == equiv_props


@pytest.mark.parametrize(
    "style,equiv",
    [
        (
            "margin: 1px; margin-top: inherit",
            "margin-bottom: 1px; margin-right: 1px; margin-left: 1px",
        ),
        ("margin-top: inherit", ""),
        ("margin-top: initial", ""),
    ],
)
def test_css_none_absent(style, equiv):
    assert_same_resolution(style, equiv)


@pytest.mark.parametrize(
    "size,resolved",
    [
        ("xx-small", "6pt"),
        ("x-small", f"{7.5:f}pt"),
        ("small", f"{9.6:f}pt"),
        ("medium", "12pt"),
        ("large", f"{13.5:f}pt"),
        ("x-large", "18pt"),
        ("xx-large", "24pt"),
        ("8px", "6pt"),
        ("1.25pc", "15pt"),
        (".25in", "18pt"),
        ("02.54cm", "72pt"),
        ("25.4mm", "72pt"),
        ("101.6q", "72pt"),
        ("101.6q", "72pt"),
    ],
)
@pytest.mark.parametrize("relative_to", [None, "16pt"])  # invariant to inherited size
def test_css_absolute_font_size(size, relative_to, resolved):
    if relative_to is None:
        inherited = None
    else:
        inherited = {"font-size": relative_to}
    assert_resolves(f"font-size: {size}", {"font-size": resolved}, inherited=inherited)


@pytest.mark.parametrize(
    "size,relative_to,resolved",
    [
        ("1em", None, "12pt"),
        ("1.0em", None, "12pt"),
        ("1.25em", None, "15pt"),
        ("1em", "16pt", "16pt"),
        ("1.0em", "16pt", "16pt"),
        ("1.25em", "16pt", "20pt"),
        ("1rem", "16pt", "12pt"),
        ("1.0rem", "16pt", "12pt"),
        ("1.25rem", "16pt", "15pt"),
        ("100%", None, "12pt"),
        ("125%", None, "15pt"),
        ("100%", "16pt", "16pt"),
        ("125%", "16pt", "20pt"),
        ("2ex", None, "12pt"),
        ("2.0ex", None, "12pt"),
        ("2.50ex", None, "15pt"),
        ("inherit", "16pt", "16pt"),
        ("smaller", None, "10pt"),
        ("smaller", "18pt", "15pt"),
        ("larger", None, f"{14.4:f}pt"),
        ("larger", "15pt", "18pt"),
    ],
)
def test_css_relative_font_size(size, relative_to, resolved):
    if relative_to is None:
        inherited = None
    else:
        inherited = {"font-size": relative_to}
    assert_resolves(f"font-size: {size}", {"font-size": resolved}, inherited=inherited)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\quantum\tests\test_density.py ===
from sympy.core.numbers import Rational
from sympy.core.singleton import S
from sympy.core.symbol import symbols
from sympy.functions.elementary.exponential import log
from sympy.external import import_module
from sympy.physics.quantum.density import Density, entropy, fidelity
from sympy.physics.quantum.state import Ket, TimeDepKet
from sympy.physics.quantum.qubit import Qubit
from sympy.physics.quantum.represent import represent
from sympy.physics.quantum.dagger import Dagger
from sympy.physics.quantum.cartesian import XKet, PxKet, PxOp, XOp
from sympy.physics.quantum.spin import JzKet
from sympy.physics.quantum.operator import OuterProduct
from sympy.physics.quantum.trace import Tr
from sympy.functions import sqrt
from sympy.testing.pytest import raises
from sympy.physics.quantum.matrixutils import scipy_sparse_matrix
from sympy.physics.quantum.tensorproduct import TensorProduct


def test_eval_args():
    # check instance created
    assert isinstance(Density([Ket(0), 0.5], [Ket(1), 0.5]), Density)
    assert isinstance(Density([Qubit('00'), 1/sqrt(2)],
                              [Qubit('11'), 1/sqrt(2)]), Density)

    #test if Qubit object type preserved
    d = Density([Qubit('00'), 1/sqrt(2)], [Qubit('11'), 1/sqrt(2)])
    for (state, prob) in d.args:
        assert isinstance(state, Qubit)

    # check for value error, when prob is not provided
    raises(ValueError, lambda: Density([Ket(0)], [Ket(1)]))


def test_doit():

    x, y = symbols('x y')
    A, B, C, D, E, F = symbols('A B C D E F', commutative=False)
    d = Density([XKet(), 0.5], [PxKet(), 0.5])
    assert (0.5*(PxKet()*Dagger(PxKet())) +
            0.5*(XKet()*Dagger(XKet()))) == d.doit()

    # check for kets with expr in them
    d_with_sym = Density([XKet(x*y), 0.5], [PxKet(x*y), 0.5])
    assert (0.5*(PxKet(x*y)*Dagger(PxKet(x*y))) +
            0.5*(XKet(x*y)*Dagger(XKet(x*y)))) == d_with_sym.doit()

    d = Density([(A + B)*C, 1.0])
    assert d.doit() == (1.0*A*C*Dagger(C)*Dagger(A) +
                        1.0*A*C*Dagger(C)*Dagger(B) +
                        1.0*B*C*Dagger(C)*Dagger(A) +
                        1.0*B*C*Dagger(C)*Dagger(B))

    #  With TensorProducts as args
    # Density with simple tensor products as args
    t = TensorProduct(A, B, C)
    d = Density([t, 1.0])
    assert d.doit() == \
        1.0 * TensorProduct(A*Dagger(A), B*Dagger(B), C*Dagger(C))

    # Density with multiple Tensorproducts as states
    t2 = TensorProduct(A, B)
    t3 = TensorProduct(C, D)

    d = Density([t2, 0.5], [t3, 0.5])
    assert d.doit() == (0.5 * TensorProduct(A*Dagger(A), B*Dagger(B)) +
                        0.5 * TensorProduct(C*Dagger(C), D*Dagger(D)))

    #Density with mixed states
    d = Density([t2 + t3, 1.0])
    assert d.doit() == (1.0 * TensorProduct(A*Dagger(A), B*Dagger(B)) +
                        1.0 * TensorProduct(A*Dagger(C), B*Dagger(D)) +
                        1.0 * TensorProduct(C*Dagger(A), D*Dagger(B)) +
                        1.0 * TensorProduct(C*Dagger(C), D*Dagger(D)))

    #Density operators with spin states
    tp1 = TensorProduct(JzKet(1, 1), JzKet(1, -1))
    d = Density([tp1, 1])

    # full trace
    t = Tr(d)
    assert t.doit() == 1

    #Partial trace on density operators with spin states
    t = Tr(d, [0])
    assert t.doit() == JzKet(1, -1) * Dagger(JzKet(1, -1))
    t = Tr(d, [1])
    assert t.doit() == JzKet(1, 1) * Dagger(JzKet(1, 1))

    # with another spin state
    tp2 = TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))
    d = Density([tp2, 1])

    #full trace
    t = Tr(d)
    assert t.doit() == 1

    #Partial trace on density operators with spin states
    t = Tr(d, [0])
    assert t.doit() == JzKet(S.Half, Rational(-1, 2)) * Dagger(JzKet(S.Half, Rational(-1, 2)))
    t = Tr(d, [1])
    assert t.doit() == JzKet(S.Half, S.Half) * Dagger(JzKet(S.Half, S.Half))


def test_apply_op():
    d = Density([Ket(0), 0.5], [Ket(1), 0.5])
    assert d.apply_op(XOp()) == Density([XOp()*Ket(0), 0.5],
                                        [XOp()*Ket(1), 0.5])


def test_represent():
    x, y = symbols('x y')
    d = Density([XKet(), 0.5], [PxKet(), 0.5])
    assert (represent(0.5*(PxKet()*Dagger(PxKet()))) +
            represent(0.5*(XKet()*Dagger(XKet())))) == represent(d)

    # check for kets with expr in them
    d_with_sym = Density([XKet(x*y), 0.5], [PxKet(x*y), 0.5])
    assert (represent(0.5*(PxKet(x*y)*Dagger(PxKet(x*y)))) +
            represent(0.5*(XKet(x*y)*Dagger(XKet(x*y))))) == \
        represent(d_with_sym)

    # check when given explicit basis
    assert (represent(0.5*(XKet()*Dagger(XKet())), basis=PxOp()) +
            represent(0.5*(PxKet()*Dagger(PxKet())), basis=PxOp())) == \
        represent(d, basis=PxOp())


def test_states():
    d = Density([Ket(0), 0.5], [Ket(1), 0.5])
    states = d.states()
    assert states[0] == Ket(0) and states[1] == Ket(1)


def test_probs():
    d = Density([Ket(0), .75], [Ket(1), 0.25])
    probs = d.probs()
    assert probs[0] == 0.75 and probs[1] == 0.25

    #probs can be symbols
    x, y = symbols('x y')
    d = Density([Ket(0), x], [Ket(1), y])
    probs = d.probs()
    assert probs[0] == x and probs[1] == y


def test_get_state():
    x, y = symbols('x y')
    d = Density([Ket(0), x], [Ket(1), y])
    states = (d.get_state(0), d.get_state(1))
    assert states[0] == Ket(0) and states[1] == Ket(1)


def test_get_prob():
    x, y = symbols('x y')
    d = Density([Ket(0), x], [Ket(1), y])
    probs = (d.get_prob(0), d.get_prob(1))
    assert probs[0] == x and probs[1] == y


def test_entropy():
    up = JzKet(S.Half, S.Half)
    down = JzKet(S.Half, Rational(-1, 2))
    d = Density((up, S.Half), (down, S.Half))

    # test for density object
    ent = entropy(d)
    assert entropy(d) == log(2)/2
    assert d.entropy() == log(2)/2

    np = import_module('numpy', min_module_version='1.4.0')
    if np:
        #do this test only if 'numpy' is available on test machine
        np_mat = represent(d, format='numpy')
        ent = entropy(np_mat)
        assert isinstance(np_mat, np.ndarray)
        assert ent.real == 0.69314718055994529
        assert ent.imag == 0

    scipy = import_module('scipy', import_kwargs={'fromlist': ['sparse']})
    if scipy and np:
        #do this test only if numpy and scipy are available
        mat = represent(d, format="scipy.sparse")
        assert isinstance(mat, scipy_sparse_matrix)
        assert ent.real == 0.69314718055994529
        assert ent.imag == 0


def test_eval_trace():
    up = JzKet(S.Half, S.Half)
    down = JzKet(S.Half, Rational(-1, 2))
    d = Density((up, 0.5), (down, 0.5))

    t = Tr(d)
    assert t.doit() == 1.0

    #test dummy time dependent states
    class TestTimeDepKet(TimeDepKet):
        def _eval_trace(self, bra, **options):
            return 1

    x, t = symbols('x t')
    k1 = TestTimeDepKet(0, 0.5)
    k2 = TestTimeDepKet(0, 1)
    d = Density([k1, 0.5], [k2, 0.5])
    assert d.doit() == (0.5 * OuterProduct(k1, k1.dual) +
                        0.5 * OuterProduct(k2, k2.dual))

    t = Tr(d)
    assert t.doit() == 1.0


def test_fidelity():
    #test with kets
    up = JzKet(S.Half, S.Half)
    down = JzKet(S.Half, Rational(-1, 2))
    updown = (S.One/sqrt(2))*up + (S.One/sqrt(2))*down

    #check with matrices
    up_dm = represent(up * Dagger(up))
    down_dm = represent(down * Dagger(down))
    updown_dm = represent(updown * Dagger(updown))

    assert abs(fidelity(up_dm, up_dm) - 1) < 1e-3
    assert fidelity(up_dm, down_dm) < 1e-3
    assert abs(fidelity(up_dm, updown_dm) - (S.One/sqrt(2))) < 1e-3
    assert abs(fidelity(updown_dm, down_dm) - (S.One/sqrt(2))) < 1e-3

    #check with density
    up_dm = Density([up, 1.0])
    down_dm = Density([down, 1.0])
    updown_dm = Density([updown, 1.0])

    assert abs(fidelity(up_dm, up_dm) - 1) < 1e-3
    assert abs(fidelity(up_dm, down_dm)) < 1e-3
    assert abs(fidelity(up_dm, updown_dm) - (S.One/sqrt(2))) < 1e-3
    assert abs(fidelity(updown_dm, down_dm) - (S.One/sqrt(2))) < 1e-3

    #check mixed states with density
    updown2 = sqrt(3)/2*up + S.Half*down
    d1 = Density([updown, 0.25], [updown2, 0.75])
    d2 = Density([updown, 0.75], [updown2, 0.25])
    assert abs(fidelity(d1, d2) - 0.991) < 1e-3
    assert abs(fidelity(d2, d1) - fidelity(d1, d2)) < 1e-3

    #using qubits/density(pure states)
    state1 = Qubit('0')
    state2 = Qubit('1')
    state3 = S.One/sqrt(2)*state1 + S.One/sqrt(2)*state2
    state4 = sqrt(Rational(2, 3))*state1 + S.One/sqrt(3)*state2

    state1_dm = Density([state1, 1])
    state2_dm = Density([state2, 1])
    state3_dm = Density([state3, 1])

    assert fidelity(state1_dm, state1_dm) == 1
    assert fidelity(state1_dm, state2_dm) == 0
    assert abs(fidelity(state1_dm, state3_dm) - 1/sqrt(2)) < 1e-3
    assert abs(fidelity(state3_dm, state2_dm) - 1/sqrt(2)) < 1e-3

    #using qubits/density(mixed states)
    d1 = Density([state3, 0.70], [state4, 0.30])
    d2 = Density([state3, 0.20], [state4, 0.80])
    assert abs(fidelity(d1, d1) - 1) < 1e-3
    assert abs(fidelity(d1, d2) - 0.996) < 1e-3
    assert abs(fidelity(d1, d2) - fidelity(d2, d1)) < 1e-3

    #TODO: test for invalid arguments
    # non-square matrix
    mat1 = [[0, 0],
            [0, 0],
            [0, 0]]

    mat2 = [[0, 0],
            [0, 0]]
    raises(ValueError, lambda: fidelity(mat1, mat2))

    # unequal dimensions
    mat1 = [[0, 0],
            [0, 0]]
    mat2 = [[0, 0, 0],
            [0, 0, 0],
            [0, 0, 0]]
    raises(ValueError, lambda: fidelity(mat1, mat2))

    # unsupported data-type
    x, y = 1, 2  # random values that is not a matrix
    raises(ValueError, lambda: fidelity(x, y))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\frame\methods\test_dropna.py ===
import datetime

import dateutil
import numpy as np
import pytest

import pandas as pd
from pandas import (
    DataFrame,
    Series,
)
import pandas._testing as tm


class TestDataFrameMissingData:
    def test_dropEmptyRows(self, float_frame):
        N = len(float_frame.index)
        mat = np.random.default_rng(2).standard_normal(N)
        mat[:5] = np.nan

        frame = DataFrame({"foo": mat}, index=float_frame.index)
        original = Series(mat, index=float_frame.index, name="foo")
        expected = original.dropna()
        inplace_frame1, inplace_frame2 = frame.copy(), frame.copy()

        smaller_frame = frame.dropna(how="all")
        # check that original was preserved
        tm.assert_series_equal(frame["foo"], original)
        return_value = inplace_frame1.dropna(how="all", inplace=True)
        tm.assert_series_equal(smaller_frame["foo"], expected)
        tm.assert_series_equal(inplace_frame1["foo"], expected)
        assert return_value is None

        smaller_frame = frame.dropna(how="all", subset=["foo"])
        return_value = inplace_frame2.dropna(how="all", subset=["foo"], inplace=True)
        tm.assert_series_equal(smaller_frame["foo"], expected)
        tm.assert_series_equal(inplace_frame2["foo"], expected)
        assert return_value is None

    def test_dropIncompleteRows(self, float_frame):
        N = len(float_frame.index)
        mat = np.random.default_rng(2).standard_normal(N)
        mat[:5] = np.nan

        frame = DataFrame({"foo": mat}, index=float_frame.index)
        frame["bar"] = 5
        original = Series(mat, index=float_frame.index, name="foo")
        inp_frame1, inp_frame2 = frame.copy(), frame.copy()

        smaller_frame = frame.dropna()
        tm.assert_series_equal(frame["foo"], original)
        return_value = inp_frame1.dropna(inplace=True)

        exp = Series(mat[5:], index=float_frame.index[5:], name="foo")
        tm.assert_series_equal(smaller_frame["foo"], exp)
        tm.assert_series_equal(inp_frame1["foo"], exp)
        assert return_value is None

        samesize_frame = frame.dropna(subset=["bar"])
        tm.assert_series_equal(frame["foo"], original)
        assert (frame["bar"] == 5).all()
        return_value = inp_frame2.dropna(subset=["bar"], inplace=True)
        tm.assert_index_equal(samesize_frame.index, float_frame.index)
        tm.assert_index_equal(inp_frame2.index, float_frame.index)
        assert return_value is None

    def test_dropna(self):
        df = DataFrame(np.random.default_rng(2).standard_normal((6, 4)))
        df.iloc[:2, 2] = np.nan

        dropped = df.dropna(axis=1)
        expected = df.loc[:, [0, 1, 3]]
        inp = df.copy()
        return_value = inp.dropna(axis=1, inplace=True)
        tm.assert_frame_equal(dropped, expected)
        tm.assert_frame_equal(inp, expected)
        assert return_value is None

        dropped = df.dropna(axis=0)
        expected = df.loc[list(range(2, 6))]
        inp = df.copy()
        return_value = inp.dropna(axis=0, inplace=True)
        tm.assert_frame_equal(dropped, expected)
        tm.assert_frame_equal(inp, expected)
        assert return_value is None

        # threshold
        dropped = df.dropna(axis=1, thresh=5)
        expected = df.loc[:, [0, 1, 3]]
        inp = df.copy()
        return_value = inp.dropna(axis=1, thresh=5, inplace=True)
        tm.assert_frame_equal(dropped, expected)
        tm.assert_frame_equal(inp, expected)
        assert return_value is None

        dropped = df.dropna(axis=0, thresh=4)
        expected = df.loc[range(2, 6)]
        inp = df.copy()
        return_value = inp.dropna(axis=0, thresh=4, inplace=True)
        tm.assert_frame_equal(dropped, expected)
        tm.assert_frame_equal(inp, expected)
        assert return_value is None

        dropped = df.dropna(axis=1, thresh=4)
        tm.assert_frame_equal(dropped, df)

        dropped = df.dropna(axis=1, thresh=3)
        tm.assert_frame_equal(dropped, df)

        # subset
        dropped = df.dropna(axis=0, subset=[0, 1, 3])
        inp = df.copy()
        return_value = inp.dropna(axis=0, subset=[0, 1, 3], inplace=True)
        tm.assert_frame_equal(dropped, df)
        tm.assert_frame_equal(inp, df)
        assert return_value is None

        # all
        dropped = df.dropna(axis=1, how="all")
        tm.assert_frame_equal(dropped, df)

        df[2] = np.nan
        dropped = df.dropna(axis=1, how="all")
        expected = df.loc[:, [0, 1, 3]]
        tm.assert_frame_equal(dropped, expected)

        # bad input
        msg = "No axis named 3 for object type DataFrame"
        with pytest.raises(ValueError, match=msg):
            df.dropna(axis=3)

    def test_drop_and_dropna_caching(self):
        # tst that cacher updates
        original = Series([1, 2, np.nan], name="A")
        expected = Series([1, 2], dtype=original.dtype, name="A")
        df = DataFrame({"A": original.values.copy()})
        df2 = df.copy()
        df["A"].dropna()
        tm.assert_series_equal(df["A"], original)

        ser = df["A"]
        return_value = ser.dropna(inplace=True)
        tm.assert_series_equal(ser, expected)
        tm.assert_series_equal(df["A"], original)
        assert return_value is None

        df2["A"].drop([1])
        tm.assert_series_equal(df2["A"], original)

        ser = df2["A"]
        return_value = ser.drop([1], inplace=True)
        tm.assert_series_equal(ser, original.drop([1]))
        tm.assert_series_equal(df2["A"], original)
        assert return_value is None

    def test_dropna_corner(self, float_frame):
        # bad input
        msg = "invalid how option: foo"
        with pytest.raises(ValueError, match=msg):
            float_frame.dropna(how="foo")
        # non-existent column - 8303
        with pytest.raises(KeyError, match=r"^\['X'\]$"):
            float_frame.dropna(subset=["A", "X"])

    def test_dropna_multiple_axes(self):
        df = DataFrame(
            [
                [1, np.nan, 2, 3],
                [4, np.nan, 5, 6],
                [np.nan, np.nan, np.nan, np.nan],
                [7, np.nan, 8, 9],
            ]
        )

        # GH20987
        with pytest.raises(TypeError, match="supplying multiple axes"):
            df.dropna(how="all", axis=[0, 1])
        with pytest.raises(TypeError, match="supplying multiple axes"):
            df.dropna(how="all", axis=(0, 1))

        inp = df.copy()
        with pytest.raises(TypeError, match="supplying multiple axes"):
            inp.dropna(how="all", axis=(0, 1), inplace=True)

    def test_dropna_tz_aware_datetime(self, using_infer_string):
        # GH13407

        df = DataFrame()
        if using_infer_string:
            df.columns = df.columns.astype("str")
        dt1 = datetime.datetime(2015, 1, 1, tzinfo=dateutil.tz.tzutc())
        dt2 = datetime.datetime(2015, 2, 2, tzinfo=dateutil.tz.tzutc())
        df["Time"] = [dt1]
        result = df.dropna(axis=0)
        expected = DataFrame({"Time": [dt1]})
        tm.assert_frame_equal(result, expected)

        # Ex2
        df = DataFrame({"Time": [dt1, None, np.nan, dt2]})
        result = df.dropna(axis=0)
        expected = DataFrame([dt1, dt2], columns=["Time"], index=[0, 3])
        tm.assert_frame_equal(result, expected)

    def test_dropna_categorical_interval_index(self):
        # GH 25087
        ii = pd.IntervalIndex.from_breaks([0, 2.78, 3.14, 6.28])
        ci = pd.CategoricalIndex(ii)
        df = DataFrame({"A": list("abc")}, index=ci)

        expected = df
        result = df.dropna()
        tm.assert_frame_equal(result, expected)

    def test_dropna_with_duplicate_columns(self):
        df = DataFrame(
            {
                "A": np.random.default_rng(2).standard_normal(5),
                "B": np.random.default_rng(2).standard_normal(5),
                "C": np.random.default_rng(2).standard_normal(5),
                "D": ["a", "b", "c", "d", "e"],
            }
        )
        df.iloc[2, [0, 1, 2]] = np.nan
        df.iloc[0, 0] = np.nan
        df.iloc[1, 1] = np.nan
        df.iloc[:, 3] = np.nan
        expected = df.dropna(subset=["A", "B", "C"], how="all")
        expected.columns = ["A", "A", "B", "C"]

        df.columns = ["A", "A", "B", "C"]

        result = df.dropna(subset=["A", "C"], how="all")
        tm.assert_frame_equal(result, expected)

    def test_set_single_column_subset(self):
        # GH 41021
        df = DataFrame({"A": [1, 2, 3], "B": list("abc"), "C": [4, np.nan, 5]})
        expected = DataFrame(
            {"A": [1, 3], "B": list("ac"), "C": [4.0, 5.0]}, index=[0, 2]
        )
        result = df.dropna(subset="C")
        tm.assert_frame_equal(result, expected)

    def test_single_column_not_present_in_axis(self):
        # GH 41021
        df = DataFrame({"A": [1, 2, 3]})

        # Column not present
        with pytest.raises(KeyError, match="['D']"):
            df.dropna(subset="D", axis=0)

    def test_subset_is_nparray(self):
        # GH 41021
        df = DataFrame({"A": [1, 2, np.nan], "B": list("abc"), "C": [4, np.nan, 5]})
        expected = DataFrame({"A": [1.0], "B": ["a"], "C": [4.0]})
        result = df.dropna(subset=np.array(["A", "C"]))
        tm.assert_frame_equal(result, expected)

    def test_no_nans_in_frame(self, axis):
        # GH#41965
        df = DataFrame([[1, 2], [3, 4]], columns=pd.RangeIndex(0, 2))
        expected = df.copy()
        result = df.dropna(axis=axis)
        tm.assert_frame_equal(result, expected, check_index_type=True)

    def test_how_thresh_param_incompatible(self):
        # GH46575
        df = DataFrame([1, 2, pd.NA])
        msg = "You cannot set both the how and thresh arguments at the same time"
        with pytest.raises(TypeError, match=msg):
            df.dropna(how="all", thresh=2)

        with pytest.raises(TypeError, match=msg):
            df.dropna(how="any", thresh=2)

        with pytest.raises(TypeError, match=msg):
            df.dropna(how=None, thresh=None)

    @pytest.mark.parametrize("val", [1, 1.5])
    def test_dropna_ignore_index(self, val):
        # GH#31725
        df = DataFrame({"a": [1, 2, val]}, index=[3, 2, 1])
        result = df.dropna(ignore_index=True)
        expected = DataFrame({"a": [1, 2, val]})
        tm.assert_frame_equal(result, expected)

        df.dropna(ignore_index=True, inplace=True)
        tm.assert_frame_equal(df, expected)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\charset_normalizer\version.py ===
"""
Expose version
"""

from __future__ import annotations

__version__ = "3.4.2"
VERSION = __version__.split(".")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\charset_normalizer\cli\__init__.py ===
from __future__ import annotations

from .__main__ import cli_detect, query_yes_no

__all__ = (
    "cli_detect",
    "query_yes_no",
)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\et_xmlfile\__init__.py ===
from .xmlfile import xmlfile

# constants
__version__ = '2.0.0'
__author__ = 'See AUTHORS.txt'
__license__ = 'MIT'
__author_email__ = 'charlie.clark@clark-consulting.eu'
__url__ = 'https://foss.heptapod.net/openpyxl/et_xmlfile'

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\markupsafe\_native.py ===
def _escape_inner(s: str, /) -> str:
    return (
        s.replace("&", "&amp;")
        .replace(">", "&gt;")
        .replace("<", "&lt;")
        .replace("'", "&#39;")
        .replace('"', "&#34;")
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_core\tests\test_nep50_promotions.py ===
"""
This file adds basic tests to test the NEP 50 style promotion compatibility
mode.  Most of these test are likely to be simply deleted again once NEP 50
is adopted in the main test suite.  A few may be moved elsewhere.
"""

import operator

import hypothesis
import pytest
from hypothesis import strategies

import numpy as np
from numpy.testing import IS_WASM, assert_array_equal


@pytest.mark.skipif(IS_WASM, reason="wasm doesn't have support for fp errors")
def test_nep50_examples():
    res = np.uint8(1) + 2
    assert res.dtype == np.uint8

    res = np.array([1], np.uint8) + np.int64(1)
    assert res.dtype == np.int64

    res = np.array([1], np.uint8) + np.array(1, dtype=np.int64)
    assert res.dtype == np.int64

    with pytest.warns(RuntimeWarning, match="overflow"):
        res = np.uint8(100) + 200
    assert res.dtype == np.uint8

    with pytest.warns(RuntimeWarning, match="overflow"):
        res = np.float32(1) + 3e100

    assert np.isinf(res)
    assert res.dtype == np.float32

    res = np.array([0.1], np.float32) == np.float64(0.1)
    assert res[0] == False

    res = np.array([0.1], np.float32) + np.float64(0.1)
    assert res.dtype == np.float64

    res = np.array([1.], np.float32) + np.int64(3)
    assert res.dtype == np.float64


@pytest.mark.parametrize("dtype", np.typecodes["AllInteger"])
def test_nep50_weak_integers(dtype):
    # Avoids warning (different code path for scalars)
    scalar_type = np.dtype(dtype).type

    maxint = int(np.iinfo(dtype).max)

    with np.errstate(over="warn"):
        with pytest.warns(RuntimeWarning):
            res = scalar_type(100) + maxint
    assert res.dtype == dtype

    # Array operations are not expected to warn, but should give the same
    # result dtype.
    res = np.array(100, dtype=dtype) + maxint
    assert res.dtype == dtype


@pytest.mark.parametrize("dtype", np.typecodes["AllFloat"])
def test_nep50_weak_integers_with_inexact(dtype):
    # Avoids warning (different code path for scalars)
    scalar_type = np.dtype(dtype).type

    too_big_int = int(np.finfo(dtype).max) * 2

    if dtype in "dDG":
        # These dtypes currently convert to Python float internally, which
        # raises an OverflowError, while the other dtypes overflow to inf.
        # NOTE: It may make sense to normalize the behavior!
        with pytest.raises(OverflowError):
            scalar_type(1) + too_big_int

        with pytest.raises(OverflowError):
            np.array(1, dtype=dtype) + too_big_int
    else:
        # NumPy uses (or used) `int -> string -> longdouble` for the
        # conversion.  But Python may refuse `str(int)` for huge ints.
        # In that case, RuntimeWarning would be correct, but conversion
        # fails earlier (seems to happen on 32bit linux, possibly only debug).
        if dtype in "gG":
            try:
                str(too_big_int)
            except ValueError:
                pytest.skip("`huge_int -> string -> longdouble` failed")

        # Otherwise, we overflow to infinity:
        with pytest.warns(RuntimeWarning):
            res = scalar_type(1) + too_big_int
        assert res.dtype == dtype
        assert res == np.inf

        with pytest.warns(RuntimeWarning):
            # We force the dtype here, since windows may otherwise pick the
            # double instead of the longdouble loop.  That leads to slightly
            # different results (conversion of the int fails as above).
            res = np.add(np.array(1, dtype=dtype), too_big_int, dtype=dtype)
        assert res.dtype == dtype
        assert res == np.inf


@pytest.mark.parametrize("op", [operator.add, operator.pow])
def test_weak_promotion_scalar_path(op):
    # Some additional paths exercising the weak scalars.

    # Integer path:
    res = op(np.uint8(3), 5)
    assert res == op(3, 5)
    assert res.dtype == np.uint8 or res.dtype == bool  # noqa: PLR1714

    with pytest.raises(OverflowError):
        op(np.uint8(3), 1000)

    # Float path:
    res = op(np.float32(3), 5.)
    assert res == op(3., 5.)
    assert res.dtype == np.float32 or res.dtype == bool  # noqa: PLR1714


def test_nep50_complex_promotion():
    with pytest.warns(RuntimeWarning, match=".*overflow"):
        res = np.complex64(3) + complex(2**300)

    assert type(res) == np.complex64


def test_nep50_integer_conversion_errors():
    # Implementation for error paths is mostly missing (as of writing)
    with pytest.raises(OverflowError, match=".*uint8"):
        np.array([1], np.uint8) + 300

    with pytest.raises(OverflowError, match=".*uint8"):
        np.uint8(1) + 300

    # Error message depends on platform (maybe unsigned int or unsigned long)
    with pytest.raises(OverflowError,
            match="Python integer -1 out of bounds for uint8"):
        np.uint8(1) + -1


def test_nep50_with_axisconcatenator():
    # Concatenate/r_ does not promote, so this has to error:
    with pytest.raises(OverflowError):
        np.r_[np.arange(5, dtype=np.int8), 255]


@pytest.mark.parametrize("ufunc", [np.add, np.power])
def test_nep50_huge_integers(ufunc):
    # Very large integers are complicated, because they go to uint64 or
    # object dtype.  This tests covers a few possible paths.
    with pytest.raises(OverflowError):
        ufunc(np.int64(0), 2**63)  # 2**63 too large for int64

    with pytest.raises(OverflowError):
        ufunc(np.uint64(0), 2**64)  # 2**64 cannot be represented by uint64

    # However, 2**63 can be represented by the uint64 (and that is used):
    res = ufunc(np.uint64(1), 2**63)

    assert res.dtype == np.uint64
    assert res == ufunc(1, 2**63, dtype=object)

    # The following paths fail to warn correctly about the change:
    with pytest.raises(OverflowError):
        ufunc(np.int64(1), 2**63)  # np.array(2**63) would go to uint

    with pytest.raises(OverflowError):
        ufunc(np.int64(1), 2**100)  # np.array(2**100) would go to object

    # This would go to object and thus a Python float, not a NumPy one:
    res = ufunc(1.0, 2**100)
    assert isinstance(res, np.float64)


def test_nep50_in_concat_and_choose():
    res = np.concatenate([np.float32(1), 1.], axis=None)
    assert res.dtype == "float32"

    res = np.choose(1, [np.float32(1), 1.])
    assert res.dtype == "float32"


@pytest.mark.parametrize("expected,dtypes,optional_dtypes", [
        (np.float32, [np.float32],
            [np.float16, 0.0, np.uint16, np.int16, np.int8, 0]),
        (np.complex64, [np.float32, 0j],
            [np.float16, 0.0, np.uint16, np.int16, np.int8, 0]),
        (np.float32, [np.int16, np.uint16, np.float16],
            [np.int8, np.uint8, np.float32, 0., 0]),
        (np.int32, [np.int16, np.uint16],
            [np.int8, np.uint8, 0, np.bool]),
        ])
@hypothesis.given(data=strategies.data())
def test_expected_promotion(expected, dtypes, optional_dtypes, data):
    # Sample randomly while ensuring "dtypes" is always present:
    optional = data.draw(strategies.lists(
            strategies.sampled_from(dtypes + optional_dtypes)))
    all_dtypes = dtypes + optional
    dtypes_sample = data.draw(strategies.permutations(all_dtypes))

    res = np.result_type(*dtypes_sample)
    assert res == expected


@pytest.mark.parametrize("sctype",
        [np.int8, np.int16, np.int32, np.int64,
         np.uint8, np.uint16, np.uint32, np.uint64])
@pytest.mark.parametrize("other_val",
        [-2 * 100, -1, 0, 9, 10, 11, 2**63, 2 * 100])
@pytest.mark.parametrize("comp",
        [operator.eq, operator.ne, operator.le, operator.lt,
         operator.ge, operator.gt])
def test_integer_comparison(sctype, other_val, comp):
    # Test that comparisons with integers (especially out-of-bound) ones
    # works correctly.
    val_obj = 10
    val = sctype(val_obj)
    # Check that the scalar behaves the same as the python int:
    assert comp(10, other_val) == comp(val, other_val)
    assert comp(val, other_val) == comp(10, other_val)
    # Except for the result type:
    assert type(comp(val, other_val)) is np.bool

    # Check that the integer array and object array behave the same:
    val_obj = np.array([10, 10], dtype=object)
    val = val_obj.astype(sctype)
    assert_array_equal(comp(val_obj, other_val), comp(val, other_val))
    assert_array_equal(comp(other_val, val_obj), comp(other_val, val))


@pytest.mark.parametrize("arr", [
    np.ones((100, 100), dtype=np.uint8)[::2],  # not trivially iterable
    np.ones(20000, dtype=">u4"),  # cast and >buffersize
    np.ones(100, dtype=">u4"),  # fast path compatible with cast
])
def test_integer_comparison_with_cast(arr):
    # Similar to above, but mainly test a few cases that cover the slow path
    # the test is limited to unsigned ints and -1 for simplicity.
    res = arr >= -1
    assert_array_equal(res, np.ones_like(arr, dtype=bool))
    res = arr < -1
    assert_array_equal(res, np.zeros_like(arr, dtype=bool))


@pytest.mark.parametrize("comp",
        [np.equal, np.not_equal, np.less_equal, np.less,
         np.greater_equal, np.greater])
def test_integer_integer_comparison(comp):
    # Test that the NumPy comparison ufuncs work with large Python integers
    assert comp(2**200, -2**200) == comp(2**200, -2**200, dtype=object)


def create_with_scalar(sctype, value):
    return sctype(value)


def create_with_array(sctype, value):
    return np.array([value], dtype=sctype)


@pytest.mark.parametrize("sctype",
        [np.int8, np.int16, np.int32, np.int64,
         np.uint8, np.uint16, np.uint32, np.uint64])
@pytest.mark.parametrize("create", [create_with_scalar, create_with_array])
def test_oob_creation(sctype, create):
    iinfo = np.iinfo(sctype)

    with pytest.raises(OverflowError):
        create(sctype, iinfo.min - 1)

    with pytest.raises(OverflowError):
        create(sctype, iinfo.max + 1)

    with pytest.raises(OverflowError):
        create(sctype, str(iinfo.min - 1))

    with pytest.raises(OverflowError):
        create(sctype, str(iinfo.max + 1))

    assert create(sctype, iinfo.min) == iinfo.min
    assert create(sctype, iinfo.max) == iinfo.max

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_typing\_shape.py ===
from collections.abc import Sequence
from typing import Any, SupportsIndex, TypeAlias

_Shape: TypeAlias = tuple[int, ...]
_AnyShape: TypeAlias = tuple[Any, ...]

# Anything that can be coerced to a shape tuple
_ShapeLike: TypeAlias = SupportsIndex | Sequence[SupportsIndex]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\tools\ort_format_model\ort_flatbuffers_py\fbs\DimensionValueType.py ===
# automatically generated by the FlatBuffers compiler, do not modify

# namespace: fbs

class DimensionValueType(object):
    UNKNOWN = 0
    VALUE = 1
    PARAM = 2

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\__init__.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
import os
import sys

sys.path.append(os.path.dirname(__file__))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\compat\abc.py ===
# Copyright (c) 2010-2024 openpyxl


try:
    from abc import ABC
except ImportError:
    from abc import ABCMeta
    ABC = ABCMeta('ABC', (object, ), {})

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\worksheet\picture.py ===
#Autogenerated schema
from openpyxl.descriptors.serialisable import Serialisable

# same as related

class SheetBackgroundPicture(Serialisable):

    tagname = "sheetBackgroundPicture"

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\api\interchange\__init__.py ===
"""
Public API for DataFrame interchange protocol.
"""

from pandas.core.interchange.dataframe_protocol import DataFrame
from pandas.core.interchange.from_dataframe import from_dataframe

__all__ = ["from_dataframe", "DataFrame"]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\computation\check.py ===
from __future__ import annotations

from pandas.compat._optional import import_optional_dependency

ne = import_optional_dependency("numexpr", errors="warn")
NUMEXPR_INSTALLED = ne is not None

__all__ = ["NUMEXPR_INSTALLED"]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\tests\copy_view\test_astype.py ===
import pickle

import numpy as np
import pytest

from pandas.compat import HAS_PYARROW
from pandas.compat.pyarrow import pa_version_under12p0
import pandas.util._test_decorators as td

import pandas as pd
from pandas import (
    DataFrame,
    Series,
    Timestamp,
    date_range,
)
import pandas._testing as tm
from pandas.tests.copy_view.util import get_array


def test_astype_single_dtype(using_copy_on_write):
    df = DataFrame({"a": [1, 2, 3], "b": [4, 5, 6], "c": 1.5})
    df_orig = df.copy()
    df2 = df.astype("float64")

    if using_copy_on_write:
        assert np.shares_memory(get_array(df2, "c"), get_array(df, "c"))
        assert not np.shares_memory(get_array(df2, "a"), get_array(df, "a"))
    else:
        assert not np.shares_memory(get_array(df2, "c"), get_array(df, "c"))
        assert not np.shares_memory(get_array(df2, "a"), get_array(df, "a"))

    # mutating df2 triggers a copy-on-write for that column/block
    df2.iloc[0, 2] = 5.5
    if using_copy_on_write:
        assert not np.shares_memory(get_array(df2, "c"), get_array(df, "c"))
    tm.assert_frame_equal(df, df_orig)

    # mutating parent also doesn't update result
    df2 = df.astype("float64")
    df.iloc[0, 2] = 5.5
    tm.assert_frame_equal(df2, df_orig.astype("float64"))


@pytest.mark.parametrize("dtype", ["int64", "Int64"])
@pytest.mark.parametrize("new_dtype", ["int64", "Int64", "int64[pyarrow]"])
def test_astype_avoids_copy(using_copy_on_write, dtype, new_dtype):
    if new_dtype == "int64[pyarrow]":
        pytest.importorskip("pyarrow")
    df = DataFrame({"a": [1, 2, 3]}, dtype=dtype)
    df_orig = df.copy()
    df2 = df.astype(new_dtype)

    if using_copy_on_write:
        assert np.shares_memory(get_array(df2, "a"), get_array(df, "a"))
    else:
        assert not np.shares_memory(get_array(df2, "a"), get_array(df, "a"))

    # mutating df2 triggers a copy-on-write for that column/block
    df2.iloc[0, 0] = 10
    if using_copy_on_write:
        assert not np.shares_memory(get_array(df2, "a"), get_array(df, "a"))
    tm.assert_frame_equal(df, df_orig)

    # mutating parent also doesn't update result
    df2 = df.astype(new_dtype)
    df.iloc[0, 0] = 100
    tm.assert_frame_equal(df2, df_orig.astype(new_dtype))


@pytest.mark.parametrize("dtype", ["float64", "int32", "Int32", "int32[pyarrow]"])
def test_astype_different_target_dtype(using_copy_on_write, dtype):
    if dtype == "int32[pyarrow]":
        pytest.importorskip("pyarrow")
    df = DataFrame({"a": [1, 2, 3]})
    df_orig = df.copy()
    df2 = df.astype(dtype)

    assert not np.shares_memory(get_array(df2, "a"), get_array(df, "a"))
    if using_copy_on_write:
        assert df2._mgr._has_no_reference(0)

    df2.iloc[0, 0] = 5
    tm.assert_frame_equal(df, df_orig)

    # mutating parent also doesn't update result
    df2 = df.astype(dtype)
    df.iloc[0, 0] = 100
    tm.assert_frame_equal(df2, df_orig.astype(dtype))


@td.skip_array_manager_invalid_test
def test_astype_numpy_to_ea():
    ser = Series([1, 2, 3])
    with pd.option_context("mode.copy_on_write", True):
        result = ser.astype("Int64")
    assert np.shares_memory(get_array(ser), get_array(result))


@pytest.mark.parametrize(
    "dtype, new_dtype", [("object", "string"), ("string", "object")]
)
def test_astype_string_and_object(using_copy_on_write, dtype, new_dtype):
    df = DataFrame({"a": ["a", "b", "c"]}, dtype=dtype)
    df_orig = df.copy()
    df2 = df.astype(new_dtype)

    if using_copy_on_write:
        assert np.shares_memory(get_array(df2, "a"), get_array(df, "a"))
    else:
        assert not np.shares_memory(get_array(df2, "a"), get_array(df, "a"))

    df2.iloc[0, 0] = "x"
    tm.assert_frame_equal(df, df_orig)


@pytest.mark.parametrize(
    "dtype, new_dtype", [("object", "string"), ("string", "object")]
)
def test_astype_string_and_object_update_original(
    using_copy_on_write, dtype, new_dtype
):
    df = DataFrame({"a": ["a", "b", "c"]}, dtype=dtype)
    df2 = df.astype(new_dtype)
    df_orig = df2.copy()

    if using_copy_on_write:
        assert np.shares_memory(get_array(df2, "a"), get_array(df, "a"))
    else:
        assert not np.shares_memory(get_array(df2, "a"), get_array(df, "a"))

    df.iloc[0, 0] = "x"
    tm.assert_frame_equal(df2, df_orig)


def test_astype_str_copy_on_pickle_roundrip():
    # TODO(infer_string) this test can be removed after 3.0 (once str is the default)
    # https://github.com/pandas-dev/pandas/issues/54654
    # ensure_string_array may alter array inplace
    base = Series(np.array([(1, 2), None, 1], dtype="object"))
    base_copy = pickle.loads(pickle.dumps(base))
    base_copy.astype(str)
    tm.assert_series_equal(base, base_copy)


def test_astype_string_copy_on_pickle_roundrip(any_string_dtype):
    # https://github.com/pandas-dev/pandas/issues/54654
    # ensure_string_array may alter array inplace
    base = Series(np.array([(1, 2), None, 1], dtype="object"))
    base_copy = pickle.loads(pickle.dumps(base))
    base_copy.astype(any_string_dtype)
    tm.assert_series_equal(base, base_copy)


def test_astype_string_read_only_on_pickle_roundrip(any_string_dtype):
    # https://github.com/pandas-dev/pandas/issues/54654
    # ensure_string_array may alter read-only array inplace
    base = Series(np.array([(1, 2), None, 1], dtype="object"))
    base_copy = pickle.loads(pickle.dumps(base))
    base_copy._values.flags.writeable = False
    base_copy.astype(any_string_dtype)
    tm.assert_series_equal(base, base_copy)


def test_astype_dict_dtypes(using_copy_on_write):
    df = DataFrame(
        {"a": [1, 2, 3], "b": [4, 5, 6], "c": Series([1.5, 1.5, 1.5], dtype="float64")}
    )
    df_orig = df.copy()
    df2 = df.astype({"a": "float64", "c": "float64"})

    if using_copy_on_write:
        assert np.shares_memory(get_array(df2, "c"), get_array(df, "c"))
        assert np.shares_memory(get_array(df2, "b"), get_array(df, "b"))
        assert not np.shares_memory(get_array(df2, "a"), get_array(df, "a"))
    else:
        assert not np.shares_memory(get_array(df2, "c"), get_array(df, "c"))
        assert not np.shares_memory(get_array(df2, "b"), get_array(df, "b"))
        assert not np.shares_memory(get_array(df2, "a"), get_array(df, "a"))

    # mutating df2 triggers a copy-on-write for that column/block
    df2.iloc[0, 2] = 5.5
    if using_copy_on_write:
        assert not np.shares_memory(get_array(df2, "c"), get_array(df, "c"))

    df2.iloc[0, 1] = 10
    if using_copy_on_write:
        assert not np.shares_memory(get_array(df2, "b"), get_array(df, "b"))
    tm.assert_frame_equal(df, df_orig)


def test_astype_different_datetime_resos(using_copy_on_write):
    df = DataFrame({"a": date_range("2019-12-31", periods=2, freq="D")})
    result = df.astype("datetime64[ms]")

    assert not np.shares_memory(get_array(df, "a"), get_array(result, "a"))
    if using_copy_on_write:
        assert result._mgr._has_no_reference(0)


def test_astype_different_timezones(using_copy_on_write):
    df = DataFrame(
        {"a": date_range("2019-12-31", periods=5, freq="D", tz="US/Pacific")}
    )
    result = df.astype("datetime64[ns, Europe/Berlin]")
    if using_copy_on_write:
        assert not result._mgr._has_no_reference(0)
        assert np.shares_memory(get_array(df, "a"), get_array(result, "a"))


def test_astype_different_timezones_different_reso(using_copy_on_write):
    df = DataFrame(
        {"a": date_range("2019-12-31", periods=5, freq="D", tz="US/Pacific")}
    )
    result = df.astype("datetime64[ms, Europe/Berlin]")
    if using_copy_on_write:
        assert result._mgr._has_no_reference(0)
        assert not np.shares_memory(get_array(df, "a"), get_array(result, "a"))


def test_astype_arrow_timestamp(using_copy_on_write):
    pytest.importorskip("pyarrow")
    df = DataFrame(
        {
            "a": [
                Timestamp("2020-01-01 01:01:01.000001"),
                Timestamp("2020-01-01 01:01:01.000001"),
            ]
        },
        dtype="M8[ns]",
    )
    result = df.astype("timestamp[ns][pyarrow]")
    if using_copy_on_write:
        assert not result._mgr._has_no_reference(0)
        if pa_version_under12p0:
            assert not np.shares_memory(
                get_array(df, "a"), get_array(result, "a")._pa_array
            )
        else:
            assert np.shares_memory(
                get_array(df, "a"), get_array(result, "a")._pa_array
            )


def test_convert_dtypes_infer_objects(using_copy_on_write):
    ser = Series(["a", "b", "c"])
    ser_orig = ser.copy()
    result = ser.convert_dtypes(
        convert_integer=False,
        convert_boolean=False,
        convert_floating=False,
        convert_string=False,
    )

    if using_copy_on_write:
        assert tm.shares_memory(get_array(ser), get_array(result))
    else:
        assert not np.shares_memory(get_array(ser), get_array(result))

    result.iloc[0] = "x"
    tm.assert_series_equal(ser, ser_orig)


def test_convert_dtypes(using_copy_on_write, using_infer_string):
    df = DataFrame({"a": ["a", "b"], "b": [1, 2], "c": [1.5, 2.5], "d": [True, False]})
    df_orig = df.copy()
    df2 = df.convert_dtypes()

    if using_copy_on_write:
        if using_infer_string and HAS_PYARROW:
            # TODO the default nullable string dtype still uses python storage
            # this should be changed to pyarrow if installed
            assert not tm.shares_memory(get_array(df2, "a"), get_array(df, "a"))
        else:
            assert tm.shares_memory(get_array(df2, "a"), get_array(df, "a"))
        assert tm.shares_memory(get_array(df2, "d"), get_array(df, "d"))
        assert tm.shares_memory(get_array(df2, "b"), get_array(df, "b"))
        assert tm.shares_memory(get_array(df2, "c"), get_array(df, "c"))
    else:
        assert not np.shares_memory(get_array(df2, "a"), get_array(df, "a"))
        assert not np.shares_memory(get_array(df2, "b"), get_array(df, "b"))
        assert not np.shares_memory(get_array(df2, "c"), get_array(df, "c"))
        assert not np.shares_memory(get_array(df2, "d"), get_array(df, "d"))

    df2.iloc[0, 0] = "x"
    df2.iloc[0, 1] = 10
    tm.assert_frame_equal(df, df_orig)