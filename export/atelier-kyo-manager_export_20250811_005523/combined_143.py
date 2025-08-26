
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\quantization\execution_providers\qnn\preprocess.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation. All rights reserved.
# Licensed under the MIT License. See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------
from __future__ import annotations

import logging
from pathlib import Path

import onnx

from ...fusions import FusionGelu, FusionLayerNormalization
from ...onnx_model import ONNXModel
from .fusion_lpnorm import FusionLpNormalization


def qnn_preprocess_model(
    model_input: str | Path | onnx.ModelProto,
    model_output: str | Path,
    fuse_layernorm: bool = False,
    save_as_external_data: bool = False,
    all_tensors_to_one_file: bool = False,
    external_data_location: str | None = None,
    external_data_size_threshold: int = 1024,
    external_data_convert_attribute: bool = False,
    inputs_to_make_channel_last: list[str] | None = None,
    outputs_to_make_channel_last: list[str] | None = None,
) -> bool:
    """
    If necessary, this method creates a new "pre-processed" model in preparation for
    quantization of a model to be used in QNN EP. Returns true if a new model was created.

    This method perfoms the following operations:
    - Fuse Erf sequence into a single Gelu node.
    - Fuse ReduceL2 sequence into a single LpNormalization node (p == 2).
    - (Optional) Fuse ReduceMean sequence into a single LayerNormalization node.

    Args:
        model_input: Path to the input model file or ModelProto.
        model_output: Path the output model file, which is only created if this method returns True.
        fuse_layernorm: True if ReduceMean sequences should be fused into LayerNormalization nodes.
            Defaults to False.
        save_as_external_data: True if output model should be saved with external data. Defaults to false.
        all_tensors_to_one_file: Effective only if save_as_external_data is true. Defaults to false.
            If true, save all tensors to one external file specified by external_data_location.
            If false, save each tensor to a file named with the tensor name.
        external_data_location: Effective only if save_as_external_data is true. Defaults to None.
            Specify the external file to which all tensors are saved. Path is relative
            to the model path. If not specified, the model's name is used.
        external_data_size_threshold: Effective only if save_as_external_data is true. Defaults to 1024.
            Tensors with a data size >= external_data_size_threshold are converted to external data.
            To convert every tensor with raw data to external data, set to 0.
        external_data_convert_attribute: Effective only if save_as_external_data is true. Defaults to false.
            If true, convert all tensors to external data.
            If false, convert only non-attribute tensors to external data.
        inputs_to_make_channel_last: List of graph input names to transpose to be "channel-last". For example,
            if "input0" originally has the shape (N, C, D1, D2, ..., Dn), the resulting model will change input0's
            shape to (N, D1, D2, ..., Dn, C) and add a transpose node after it.

            Original:
                input0 (N, C, D1, D2, ..., Dn) --> <Nodes>

            Updated:
                input0 (N, D1, D2, ..., Dn, C) --> Transpose --> input0_chanfirst (N, C, D1, D2, ..., Dn) --> <Nodes>

            This can potentially improve inference latency for QDQ models running on QNN EP because the
            additional transpose node may allow other transpose nodes inserted during ORT layout transformation
            to cancel out.
        outputs_to_make_channel_last: List of graph output names to transpose to be "channel-last". For example,
            if "output0" originally has the shape (N, C, D1, D2, ..., Dn), the resulting model will change output0's
            shape to (N, D1, D2, ..., Dn, C) and add a transpose node before it.

            Original:
                <Nodes> --> output0 (N, C, D1, D2, ..., Dn)

            Updated:
                <Nodes> --> output0_chanfirst (N, C, D1, D2, ..., Dn) --> Transpose --> output0 (N, D1, D2, ..., Dn, C)

            This can potentially improve inference latency for QDQ models running on QNN EP because the
            additional transpose node may allow other transpose nodes inserted during ORT layout transformation
            to cancel out.
    """
    modified = False
    model = model_input if isinstance(model_input, onnx.ModelProto) else onnx.load_model(model_input)
    onnx_model = ONNXModel(model)

    # Fuse Erf sequence into a single Gelu
    fusion_gelu = FusionGelu(onnx_model)
    if fusion_gelu.apply():
        modified = True

    # Fuse ReduceL2 sequence into a single LpNormalization node with p == 2.
    fusion_lpnorm = FusionLpNormalization(onnx_model)
    if fusion_lpnorm.apply():
        modified = True

    # Optionally, fuse ReduceMean sequence into a single LayerNormalization node.
    if fuse_layernorm:
        onnx_opset = next(x for x in model.opset_import if x.domain == "" or x.domain == "ai.onnx")

        # Need opset >= 17 to use LayerNormalization.
        if onnx_opset.version < 17:
            logging.warning(
                "Unable to fuse ReduceMean sequence into a LayerNormalization node. "
                "ONNX model must use an opset >= 17 in order to use LayerNormalization, "
                f"but found version {onnx_opset.version}. Please use onnx.version_converter to update your model."
            )
        else:
            fusion_layernorm = FusionLayerNormalization(onnx_model)
            if fusion_layernorm.apply():
                modified = True

    # Optionally, transpose inputs and/or outputs to make them "channel-last".
    if inputs_to_make_channel_last or outputs_to_make_channel_last:
        transpose_node_prefix = "Transpose_channel_"
        transpose_node_suffix: int = onnx_model.get_largest_node_name_suffix(transpose_node_prefix) + 1
        update_io_to_channel_last(
            onnx_model.model,
            inputs_to_make_channel_last,
            outputs_to_make_channel_last,
            transpose_node_name_prefix=transpose_node_prefix,
            transpose_node_name_start_suffix=transpose_node_suffix,
        )
        modified = True

    # Make sure all nodes have a name.
    unnamed_node_prefix = "qnn_preproc_node_"
    available_suffix = onnx_model.get_largest_node_name_suffix(unnamed_node_prefix) + 1
    for node in onnx_model.model.graph.node:
        if node.op_type != "Constant" and not node.name:
            new_node_name = f"{unnamed_node_prefix}{available_suffix!s}"
            available_suffix += 1
            node.name = new_node_name
            modified = True
            logging.warning(f"Node of type {node.op_type} does not have a name. Renamed to {new_node_name}.")

    if modified:
        onnx_model.topological_sort()
        onnx.save_model(
            model,
            model_output,
            save_as_external_data=save_as_external_data,
            all_tensors_to_one_file=all_tensors_to_one_file,
            location=external_data_location,
            size_threshold=external_data_size_threshold,
            convert_attribute=external_data_convert_attribute,
        )

    return modified


class InputOutputNameMap:
    def __init__(
        self,
        orig_tensor_names: set[str],
        orig_graph_inputs: dict[str, onnx.ValueInfoProto],
        orig_graph_outputs: dict[str, onnx.ValueInfoProto],
    ):
        self.orig_tensor_names = orig_tensor_names
        self.orig_graph_inputs = orig_graph_inputs
        self.orig_graph_outputs = orig_graph_outputs
        self.updated_io_names = {}
        self.new_value_infos = []

    def get_new_name(self, orig_name: str):
        if orig_name in self.updated_io_names:
            return self.updated_io_names[orig_name]

        # Make a new tensor name that is unique among all tensors in the graph.
        prefix: str = f"{orig_name}_channel_first_"
        suffix: int = -1
        for tensor_name in self.orig_tensor_names:
            if tensor_name.startswith(prefix) and tensor_name[len(prefix) :].isdigit():
                index = int(tensor_name[len(prefix) :])
                suffix = max(suffix, index)

        suffix += 1  # This is the first available suffix.
        new_name = f"{prefix}{suffix!s}"

        # Add new value_info objects for these new tensors.
        orig_value_info = self.orig_graph_inputs.get(orig_name) or self.orig_graph_outputs[orig_name]
        value_info_proto = onnx.ValueInfoProto()
        value_info_proto.CopyFrom(orig_value_info)
        value_info_proto.name = new_name
        self.new_value_infos.append(value_info_proto)

        self.updated_io_names[orig_name] = new_name
        return self.updated_io_names[orig_name]


def update_io_to_channel_last(
    model: onnx.ModelProto,
    inputs_to_update: list[str] | None,
    outputs_to_update: list[str] | None,
    transpose_node_name_prefix: str = "Transpose_channel_",
    transpose_node_name_start_suffix: int = 0,
):
    inputs_to_update = set(inputs_to_update or [])
    outputs_to_update = set(outputs_to_update or [])

    if not inputs_to_update and not outputs_to_update:
        return

    graph = model.graph
    orig_graph_inputs = {ginput.name: ginput for ginput in graph.input}
    orig_graph_outputs = {goutput.name: goutput for goutput in graph.output}

    # Check that the user passed in actual input and output names.
    for input_name in inputs_to_update:
        if input_name not in orig_graph_inputs:
            raise ValueError(f"{input_name} is not a graph input")

    for output_name in outputs_to_update:
        if output_name not in orig_graph_outputs:
            raise ValueError(f"{output_name} is not a graph output")

    orig_tensor_names = set()
    orig_tensor_names.update(set(orig_graph_inputs))
    orig_tensor_names.update(set(orig_graph_outputs))
    orig_tensor_names.update(input_name for node in graph.node for input_name in node.input if input_name)

    # Maps original input (or output) name to its updated name used within the graph.
    io_map = InputOutputNameMap(orig_tensor_names, orig_graph_inputs, orig_graph_outputs)

    # Update each node's inputs/outputs to use the transposed versions.
    for node in graph.node:
        for i in range(len(node.input)):
            if node.input[i] and node.input[i] in inputs_to_update:
                node.input[i] = io_map.get_new_name(node.input[i])
            elif node.input[i] and node.input[i] in outputs_to_update:
                node.input[i] = io_map.get_new_name(node.input[i])

        for i in range(len(node.output)):
            if node.output[i] in outputs_to_update:
                node.output[i] = io_map.get_new_name(node.output[i])

    # Update graph inputs to channel-last and a Transpose (to channel-first) after each.
    for g_input_name in inputs_to_update:
        g_input = orig_graph_inputs[g_input_name]

        if not g_input.type.HasField("tensor_type") or not g_input.type.tensor_type.HasField("shape"):
            raise ValueError(f"Expected input {g_input.name} to have a tensor_type with a shape")

        input_shape = g_input.type.tensor_type.shape
        input_rank = len(input_shape.dim)

        if input_rank < 3:
            raise ValueError(f"Expected input {g_input.name} to be of rank >= 3")

        channel_dim = onnx.TensorShapeProto.Dimension()
        channel_dim.CopyFrom(input_shape.dim[1])
        for i in range(1, input_rank - 1):
            input_shape.dim[i].CopyFrom(input_shape.dim[i + 1])
        input_shape.dim[input_rank - 1].CopyFrom(channel_dim)

        transpose_perm = list(range(input_rank))
        for i in range(input_rank):
            transpose_perm[i] = i if i < 1 else i - 1
        transpose_perm[1] = input_rank - 1

        transpose_node = onnx.helper.make_node(
            "Transpose",
            name=f"{transpose_node_name_prefix}{transpose_node_name_start_suffix!s}",
            inputs=[g_input.name],
            outputs=[io_map.get_new_name(g_input.name)],
            perm=transpose_perm,
        )
        transpose_node_name_start_suffix += 1

        graph.node.extend([transpose_node])

    # Update graph outputs to channel-last and a Transpose (from channel-first) before each.
    for g_output_name in outputs_to_update:
        g_output = orig_graph_outputs[g_output_name]
        if not g_output.type.HasField("tensor_type") or not g_output.type.tensor_type.HasField("shape"):
            raise ValueError(f"Expected output {g_output.name} to have a tensor_type with a shape")

        output_shape = g_output.type.tensor_type.shape
        output_rank = len(output_shape.dim)

        if output_rank < 3:
            raise ValueError(f"Expected output {g_output.name} to be of rank >= 3")

        channel_dim = onnx.TensorShapeProto.Dimension()
        channel_dim.CopyFrom(output_shape.dim[1])
        for i in range(1, output_rank - 1):
            output_shape.dim[i].CopyFrom(output_shape.dim[i + 1])
        output_shape.dim[output_rank - 1].CopyFrom(channel_dim)

        transpose_perm = list(range(output_rank))
        for i in range(output_rank):
            transpose_perm[i] = i if i == 0 else i + 1
        transpose_perm[output_rank - 1] = 1

        transpose_node = onnx.helper.make_node(
            "Transpose",
            name=f"{transpose_node_name_prefix}{transpose_node_name_start_suffix!s}",
            inputs=[io_map.get_new_name(g_output.name)],
            outputs=[g_output.name],
            perm=transpose_perm,
        )
        transpose_node_name_start_suffix += 1

        graph.node.extend([transpose_node])

    graph.value_info.extend(io_map.new_value_infos)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\proxy.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.
"""The Proxy implementation."""


class ProxyTypeFactory:
    """Factory for proxy types."""

    @staticmethod
    def make(ff_value, string):
        return {"ff_value": ff_value, "string": string}


class ProxyType:
    """Set of possible types of proxy.

    Each proxy type has 2 properties:    'ff_value' is value of Firefox
    profile preference,    'string' is id of proxy type.
    """

    DIRECT = ProxyTypeFactory.make(0, "DIRECT")  # Direct connection, no proxy (default on Windows).
    MANUAL = ProxyTypeFactory.make(1, "MANUAL")  # Manual proxy settings (e.g., for httpProxy).
    PAC = ProxyTypeFactory.make(2, "PAC")  # Proxy autoconfiguration from URL.
    RESERVED_1 = ProxyTypeFactory.make(3, "RESERVED1")  # Never used.
    AUTODETECT = ProxyTypeFactory.make(4, "AUTODETECT")  # Proxy autodetection (presumably with WPAD).
    SYSTEM = ProxyTypeFactory.make(5, "SYSTEM")  # Use system settings (default on Linux).
    UNSPECIFIED = ProxyTypeFactory.make(6, "UNSPECIFIED")  # Not initialized (for internal use).

    @classmethod
    def load(cls, value):
        if isinstance(value, dict) and "string" in value:
            value = value["string"]
        value = str(value).upper()
        for attr in dir(cls):
            attr_value = getattr(cls, attr)
            if isinstance(attr_value, dict) and "string" in attr_value and attr_value["string"] == value:
                return attr_value
        raise Exception(f"No proxy type is found for {value}")


class _ProxyTypeDescriptor:
    def __init__(self, name, p_type):
        self.name = name
        self.p_type = p_type

    def __get__(self, obj, cls):
        return getattr(obj, self.name)

    def __set__(self, obj, value):
        if self.name == "autodetect" and not isinstance(value, bool):
            raise ValueError("Autodetect proxy value needs to be a boolean")
        getattr(obj, "_verify_proxy_type_compatibility")(self.p_type)
        setattr(obj, "proxyType", self.p_type)
        setattr(obj, self.name, value)


class Proxy:
    """Proxy contains information about proxy type and necessary proxy
    settings."""

    proxyType = ProxyType.UNSPECIFIED
    autodetect = False
    ftpProxy = ""
    httpProxy = ""
    noProxy = ""
    proxyAutoconfigUrl = ""
    sslProxy = ""
    socksProxy = ""
    socksUsername = ""
    socksPassword = ""
    socksVersion = None

    # create descriptor type objects
    auto_detect = _ProxyTypeDescriptor("autodetect", ProxyType.AUTODETECT)
    """Gets and Sets `auto_detect`

    Usage:
    ------
    - Get
        - `self.auto_detect`
    - Set
        - `self.auto_detect` = `value`

    Parameters:
    -----------
    `value`: `str`
    """

    ftp_proxy = _ProxyTypeDescriptor("ftpProxy", ProxyType.MANUAL)
    """Gets and Sets `ftp_proxy`

    Usage:
    ------
    - Get
        - `self.ftp_proxy`
    - Set
        - `self.ftp_proxy` = `value`

    Parameters:
    -----------
    `value`: `str`
    """

    http_proxy = _ProxyTypeDescriptor("httpProxy", ProxyType.MANUAL)
    """Gets and Sets `http_proxy`

    Usage:
    ------
    - Get
        - `self.http_proxy`
    - Set
        - `self.http_proxy` = `value`

    Parameters:
    -----------
    `value`: `str`
    """

    no_proxy = _ProxyTypeDescriptor("noProxy", ProxyType.MANUAL)
    """Gets and Sets `no_proxy`

    Usage:
    ------
    - Get
        - `self.no_proxy`
    - Set
        - `self.no_proxy` = `value`

    Parameters:
    -----------
    `value`: `str`
    """

    proxy_autoconfig_url = _ProxyTypeDescriptor("proxyAutoconfigUrl", ProxyType.PAC)
    """Gets and Sets `proxy_autoconfig_url`

    Usage:
    ------
    - Get
        - `self.proxy_autoconfig_url`
    - Set
        - `self.proxy_autoconfig_url` = `value`

    Parameters:
    -----------
    `value`: `str`
    """

    ssl_proxy = _ProxyTypeDescriptor("sslProxy", ProxyType.MANUAL)
    """Gets and Sets `ssl_proxy`

    Usage:
    ------
    - Get
        - `self.ssl_proxy`
    - Set
        - `self.ssl_proxy` = `value`

    Parameters:
    -----------
    `value`: `str`
    """

    socks_proxy = _ProxyTypeDescriptor("socksProxy", ProxyType.MANUAL)
    """Gets and Sets `socks_proxy`

    Usage:
    ------
    - Get
        - `self.sock_proxy`
    - Set
        - `self.socks_proxy` = `value`

    Parameters:
    -----------
    `value`: `str`
    """

    socks_username = _ProxyTypeDescriptor("socksUsername", ProxyType.MANUAL)
    """Gets and Sets `socks_password`

    Usage:
    ------
    - Get
        - `self.socks_password`
    - Set
        - `self.socks_password` = `value`

    Parameters:
    -----------
    `value`: `str`
    """

    socks_password = _ProxyTypeDescriptor("socksPassword", ProxyType.MANUAL)
    """Gets and Sets `socks_password`

    Usage:
    ------
    - Get
        - `self.socks_password`
    - Set
        - `self.socks_password` = `value`

    Parameters:
    -----------
    `value`: `str`
    """

    socks_version = _ProxyTypeDescriptor("socksVersion", ProxyType.MANUAL)
    """Gets and Sets `socks_version`

    Usage:
    ------
    - Get
        - `self.socks_version`
    - Set
        - `self.socks_version` = `value`

    Parameters:
    -----------
    `value`: `str`
    """

    def __init__(self, raw=None):
        """Creates a new Proxy.

        :Args:
         - raw: raw proxy data. If None, default class values are used.
        """
        if raw:
            if "proxyType" in raw and raw["proxyType"]:
                self.proxy_type = ProxyType.load(raw["proxyType"])
            if "ftpProxy" in raw and raw["ftpProxy"]:
                self.ftp_proxy = raw["ftpProxy"]
            if "httpProxy" in raw and raw["httpProxy"]:
                self.http_proxy = raw["httpProxy"]
            if "noProxy" in raw and raw["noProxy"]:
                self.no_proxy = raw["noProxy"]
            if "proxyAutoconfigUrl" in raw and raw["proxyAutoconfigUrl"]:
                self.proxy_autoconfig_url = raw["proxyAutoconfigUrl"]
            if "sslProxy" in raw and raw["sslProxy"]:
                self.sslProxy = raw["sslProxy"]
            if "autodetect" in raw and raw["autodetect"]:
                self.auto_detect = raw["autodetect"]
            if "socksProxy" in raw and raw["socksProxy"]:
                self.socks_proxy = raw["socksProxy"]
            if "socksUsername" in raw and raw["socksUsername"]:
                self.socks_username = raw["socksUsername"]
            if "socksPassword" in raw and raw["socksPassword"]:
                self.socks_password = raw["socksPassword"]
            if "socksVersion" in raw and raw["socksVersion"]:
                self.socks_version = raw["socksVersion"]

    @property
    def proxy_type(self):
        """Returns proxy type as `ProxyType`."""
        return self.proxyType

    @proxy_type.setter
    def proxy_type(self, value) -> None:
        """Sets proxy type.

        :Args:
         - value: The proxy type.
        """
        self._verify_proxy_type_compatibility(value)
        self.proxyType = value

    def _verify_proxy_type_compatibility(self, compatible_proxy):
        if self.proxyType not in (ProxyType.UNSPECIFIED, compatible_proxy):
            raise ValueError(
                f"Specified proxy type ({compatible_proxy}) not compatible with current setting ({self.proxyType})"
            )

    def to_capabilities(self):
        proxy_caps = {"proxyType": self.proxyType["string"].lower()}
        proxies = [
            "autodetect",
            "ftpProxy",
            "httpProxy",
            "proxyAutoconfigUrl",
            "sslProxy",
            "noProxy",
            "socksProxy",
            "socksUsername",
            "socksPassword",
            "socksVersion",
        ]
        for proxy in proxies:
            attr_value = getattr(self, proxy)
            if attr_value:
                proxy_caps[proxy] = attr_value
        return proxy_caps

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\cffi\verifier.py ===
#
# DEPRECATED: implementation for ffi.verify()
#
import sys, os, binascii, shutil, io
from . import __version_verifier_modules__
from . import ffiplatform
from .error import VerificationError

if sys.version_info >= (3, 3):
    import importlib.machinery
    def _extension_suffixes():
        return importlib.machinery.EXTENSION_SUFFIXES[:]
else:
    import imp
    def _extension_suffixes():
        return [suffix for suffix, _, type in imp.get_suffixes()
                if type == imp.C_EXTENSION]


if sys.version_info >= (3,):
    NativeIO = io.StringIO
else:
    class NativeIO(io.BytesIO):
        def write(self, s):
            if isinstance(s, unicode):
                s = s.encode('ascii')
            super(NativeIO, self).write(s)


class Verifier(object):

    def __init__(self, ffi, preamble, tmpdir=None, modulename=None,
                 ext_package=None, tag='', force_generic_engine=False,
                 source_extension='.c', flags=None, relative_to=None, **kwds):
        if ffi._parser._uses_new_feature:
            raise VerificationError(
                "feature not supported with ffi.verify(), but only "
                "with ffi.set_source(): %s" % (ffi._parser._uses_new_feature,))
        self.ffi = ffi
        self.preamble = preamble
        if not modulename:
            flattened_kwds = ffiplatform.flatten(kwds)
        vengine_class = _locate_engine_class(ffi, force_generic_engine)
        self._vengine = vengine_class(self)
        self._vengine.patch_extension_kwds(kwds)
        self.flags = flags
        self.kwds = self.make_relative_to(kwds, relative_to)
        #
        if modulename:
            if tag:
                raise TypeError("can't specify both 'modulename' and 'tag'")
        else:
            key = '\x00'.join(['%d.%d' % sys.version_info[:2],
                               __version_verifier_modules__,
                               preamble, flattened_kwds] +
                              ffi._cdefsources)
            if sys.version_info >= (3,):
                key = key.encode('utf-8')
            k1 = hex(binascii.crc32(key[0::2]) & 0xffffffff)
            k1 = k1.lstrip('0x').rstrip('L')
            k2 = hex(binascii.crc32(key[1::2]) & 0xffffffff)
            k2 = k2.lstrip('0').rstrip('L')
            modulename = '_cffi_%s_%s%s%s' % (tag, self._vengine._class_key,
                                              k1, k2)
        suffix = _get_so_suffixes()[0]
        self.tmpdir = tmpdir or _caller_dir_pycache()
        self.sourcefilename = os.path.join(self.tmpdir, modulename + source_extension)
        self.modulefilename = os.path.join(self.tmpdir, modulename + suffix)
        self.ext_package = ext_package
        self._has_source = False
        self._has_module = False

    def write_source(self, file=None):
        """Write the C source code.  It is produced in 'self.sourcefilename',
        which can be tweaked beforehand."""
        with self.ffi._lock:
            if self._has_source and file is None:
                raise VerificationError(
                    "source code already written")
            self._write_source(file)

    def compile_module(self):
        """Write the C source code (if not done already) and compile it.
        This produces a dynamic link library in 'self.modulefilename'."""
        with self.ffi._lock:
            if self._has_module:
                raise VerificationError("module already compiled")
            if not self._has_source:
                self._write_source()
            self._compile_module()

    def load_library(self):
        """Get a C module from this Verifier instance.
        Returns an instance of a FFILibrary class that behaves like the
        objects returned by ffi.dlopen(), but that delegates all
        operations to the C module.  If necessary, the C code is written
        and compiled first.
        """
        with self.ffi._lock:
            if not self._has_module:
                self._locate_module()
                if not self._has_module:
                    if not self._has_source:
                        self._write_source()
                    self._compile_module()
            return self._load_library()

    def get_module_name(self):
        basename = os.path.basename(self.modulefilename)
        # kill both the .so extension and the other .'s, as introduced
        # by Python 3: 'basename.cpython-33m.so'
        basename = basename.split('.', 1)[0]
        # and the _d added in Python 2 debug builds --- but try to be
        # conservative and not kill a legitimate _d
        if basename.endswith('_d') and hasattr(sys, 'gettotalrefcount'):
            basename = basename[:-2]
        return basename

    def get_extension(self):
        if not self._has_source:
            with self.ffi._lock:
                if not self._has_source:
                    self._write_source()
        sourcename = ffiplatform.maybe_relative_path(self.sourcefilename)
        modname = self.get_module_name()
        return ffiplatform.get_extension(sourcename, modname, **self.kwds)

    def generates_python_module(self):
        return self._vengine._gen_python_module

    def make_relative_to(self, kwds, relative_to):
        if relative_to and os.path.dirname(relative_to):
            dirname = os.path.dirname(relative_to)
            kwds = kwds.copy()
            for key in ffiplatform.LIST_OF_FILE_NAMES:
                if key in kwds:
                    lst = kwds[key]
                    if not isinstance(lst, (list, tuple)):
                        raise TypeError("keyword '%s' should be a list or tuple"
                                        % (key,))
                    lst = [os.path.join(dirname, fn) for fn in lst]
                    kwds[key] = lst
        return kwds

    # ----------

    def _locate_module(self):
        if not os.path.isfile(self.modulefilename):
            if self.ext_package:
                try:
                    pkg = __import__(self.ext_package, None, None, ['__doc__'])
                except ImportError:
                    return      # cannot import the package itself, give up
                    # (e.g. it might be called differently before installation)
                path = pkg.__path__
            else:
                path = None
            filename = self._vengine.find_module(self.get_module_name(), path,
                                                 _get_so_suffixes())
            if filename is None:
                return
            self.modulefilename = filename
        self._vengine.collect_types()
        self._has_module = True

    def _write_source_to(self, file):
        self._vengine._f = file
        try:
            self._vengine.write_source_to_f()
        finally:
            del self._vengine._f

    def _write_source(self, file=None):
        if file is not None:
            self._write_source_to(file)
        else:
            # Write our source file to an in memory file.
            f = NativeIO()
            self._write_source_to(f)
            source_data = f.getvalue()

            # Determine if this matches the current file
            if os.path.exists(self.sourcefilename):
                with open(self.sourcefilename, "r") as fp:
                    needs_written = not (fp.read() == source_data)
            else:
                needs_written = True

            # Actually write the file out if it doesn't match
            if needs_written:
                _ensure_dir(self.sourcefilename)
                with open(self.sourcefilename, "w") as fp:
                    fp.write(source_data)

            # Set this flag
            self._has_source = True

    def _compile_module(self):
        # compile this C source
        tmpdir = os.path.dirname(self.sourcefilename)
        outputfilename = ffiplatform.compile(tmpdir, self.get_extension())
        try:
            same = ffiplatform.samefile(outputfilename, self.modulefilename)
        except OSError:
            same = False
        if not same:
            _ensure_dir(self.modulefilename)
            shutil.move(outputfilename, self.modulefilename)
        self._has_module = True

    def _load_library(self):
        assert self._has_module
        if self.flags is not None:
            return self._vengine.load_library(self.flags)
        else:
            return self._vengine.load_library()

# ____________________________________________________________

_FORCE_GENERIC_ENGINE = False      # for tests

def _locate_engine_class(ffi, force_generic_engine):
    if _FORCE_GENERIC_ENGINE:
        force_generic_engine = True
    if not force_generic_engine:
        if '__pypy__' in sys.builtin_module_names:
            force_generic_engine = True
        else:
            try:
                import _cffi_backend
            except ImportError:
                _cffi_backend = '?'
            if ffi._backend is not _cffi_backend:
                force_generic_engine = True
    if force_generic_engine:
        from . import vengine_gen
        return vengine_gen.VGenericEngine
    else:
        from . import vengine_cpy
        return vengine_cpy.VCPythonEngine

# ____________________________________________________________

_TMPDIR = None

def _caller_dir_pycache():
    if _TMPDIR:
        return _TMPDIR
    result = os.environ.get('CFFI_TMPDIR')
    if result:
        return result
    filename = sys._getframe(2).f_code.co_filename
    return os.path.abspath(os.path.join(os.path.dirname(filename),
                           '__pycache__'))

def set_tmpdir(dirname):
    """Set the temporary directory to use instead of __pycache__."""
    global _TMPDIR
    _TMPDIR = dirname

def cleanup_tmpdir(tmpdir=None, keep_so=False):
    """Clean up the temporary directory by removing all files in it
    called `_cffi_*.{c,so}` as well as the `build` subdirectory."""
    tmpdir = tmpdir or _caller_dir_pycache()
    try:
        filelist = os.listdir(tmpdir)
    except OSError:
        return
    if keep_so:
        suffix = '.c'   # only remove .c files
    else:
        suffix = _get_so_suffixes()[0].lower()
    for fn in filelist:
        if fn.lower().startswith('_cffi_') and (
                fn.lower().endswith(suffix) or fn.lower().endswith('.c')):
            try:
                os.unlink(os.path.join(tmpdir, fn))
            except OSError:
                pass
    clean_dir = [os.path.join(tmpdir, 'build')]
    for dir in clean_dir:
        try:
            for fn in os.listdir(dir):
                fn = os.path.join(dir, fn)
                if os.path.isdir(fn):
                    clean_dir.append(fn)
                else:
                    os.unlink(fn)
        except OSError:
            pass

def _get_so_suffixes():
    suffixes = _extension_suffixes()
    if not suffixes:
        # bah, no C_EXTENSION available.  Occurs on pypy without cpyext
        if sys.platform == 'win32':
            suffixes = [".pyd"]
        else:
            suffixes = [".so"]

    return suffixes

def _ensure_dir(filename):
    dirname = os.path.dirname(filename)
    if dirname and not os.path.isdir(dirname):
        os.makedirs(dirname)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\worksheet\dimensions.py ===
# Copyright (c) 2010-2024 openpyxl

from copy import copy

from openpyxl.compat import safe_string
from openpyxl.utils import (
    get_column_letter,
    get_column_interval,
    column_index_from_string,
    range_boundaries,
)
from openpyxl.utils.units import DEFAULT_COLUMN_WIDTH
from openpyxl.descriptors import (
    Integer,
    Float,
    Bool,
    Strict,
    String,
    Alias,
)
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.styles.styleable import StyleableObject
from openpyxl.utils.bound_dictionary import BoundDictionary
from openpyxl.xml.functions import Element


class Dimension(Strict, StyleableObject):
    """Information about the display properties of a row or column."""
    __fields__ = ('hidden',
                 'outlineLevel',
                 'collapsed',)

    index = Integer()
    hidden = Bool()
    outlineLevel = Integer(allow_none=True)
    outline_level = Alias('outlineLevel')
    collapsed = Bool()
    style = Alias('style_id')


    def __init__(self, index, hidden, outlineLevel,
                 collapsed, worksheet, visible=True, style=None):
        super().__init__(sheet=worksheet, style_array=style)
        self.index = index
        self.hidden = hidden
        self.outlineLevel = outlineLevel
        self.collapsed = collapsed


    def __iter__(self):
        for key in self.__fields__:
            value = getattr(self, key, None)
            if value:
                yield key, safe_string(value)


    def __copy__(self):
        cp = self.__new__(self.__class__)
        attrib = self.__dict__
        attrib['worksheet'] = self.parent
        cp.__init__(**attrib)
        cp._style = copy(self._style)
        return cp


    def __repr__(self):
        return f"<{self.__class__.__name__} Instance, Attributes={dict(self)}>"


class RowDimension(Dimension):
    """Information about the display properties of a row."""

    __fields__ = Dimension.__fields__ + ('ht', 'customFormat', 'customHeight', 's',
                                         'thickBot', 'thickTop')
    r = Alias('index')
    s = Alias('style_id')
    ht = Float(allow_none=True)
    height = Alias('ht')
    thickBot = Bool()
    thickTop = Bool()

    def __init__(self,
                 worksheet,
                 index=0,
                 ht=None,
                 customHeight=None, # do not write
                 s=None,
                 customFormat=None, # do not write
                 hidden=False,
                 outlineLevel=0,
                 outline_level=None,
                 collapsed=False,
                 visible=None,
                 height=None,
                 r=None,
                 spans=None,
                 thickBot=None,
                 thickTop=None,
                 **kw
                 ):
        if r is not None:
            index = r
        if height is not None:
            ht = height
        self.ht = ht
        if visible is not None:
            hidden = not visible
        if outline_level is not None:
            outlineLevel = outline_level
        self.thickBot = thickBot
        self.thickTop = thickTop
        super().__init__(index, hidden, outlineLevel,
                                           collapsed, worksheet, style=s)

    @property
    def customFormat(self):
        """Always true if there is a style for the row"""
        return self.has_style

    @property
    def customHeight(self):
        """Always true if there is a height for the row"""
        return self.ht is not None


class ColumnDimension(Dimension):
    """Information about the display properties of a column."""

    width = Float()
    bestFit = Bool()
    auto_size = Alias('bestFit')
    index = String()
    min = Integer(allow_none=True)
    max = Integer(allow_none=True)
    collapsed = Bool()

    __fields__ = Dimension.__fields__ + ('width', 'bestFit', 'customWidth', 'style',
                                         'min', 'max')

    def __init__(self,
                 worksheet,
                 index='A',
                 width=DEFAULT_COLUMN_WIDTH,
                 bestFit=False,
                 hidden=False,
                 outlineLevel=0,
                 outline_level=None,
                 collapsed=False,
                 style=None,
                 min=None,
                 max=None,
                 customWidth=False, # do not write
                 visible=None,
                 auto_size=None,):
        self.width = width
        self.min = min
        self.max = max
        if visible is not None:
            hidden = not visible
        if auto_size is not None:
            bestFit = auto_size
        self.bestFit = bestFit
        if outline_level is not None:
            outlineLevel = outline_level
        self.collapsed = collapsed
        super().__init__(index, hidden, outlineLevel,
                                              collapsed, worksheet, style=style)


    @property
    def customWidth(self):
        """Always true if there is a width for the column"""
        return bool(self.width)


    def reindex(self):
        """
        Set boundaries for column definition
        """
        if not all([self.min, self.max]):
            self.min = self.max = column_index_from_string(self.index)

    @property
    def range(self):
        """Return the range of cells actually covered"""
        return f"{get_column_letter(self.min)}:{get_column_letter(self.max)}"


    def to_tree(self):
        attrs = dict(self)
        if attrs.keys() != {'min', 'max'}:
            return Element("col", **attrs)


class DimensionHolder(BoundDictionary):
    """
    Allow columns to be grouped
    """

    def __init__(self, worksheet, reference="index", default_factory=None):
        self.worksheet = worksheet
        self.max_outline = None
        self.default_factory = default_factory
        super().__init__(reference, default_factory)


    def group(self, start, end=None, outline_level=1, hidden=False):
        """allow grouping a range of consecutive rows or columns together

        :param start: first row or column to be grouped (mandatory)
        :param end: last row or column to be grouped (optional, default to start)
        :param outline_level: outline level
        :param hidden: should the group be hidden on workbook open or not
        """
        if end is None:
            end = start

        if isinstance(self.default_factory(), ColumnDimension):
            new_dim = self[start]
            new_dim.outline_level = outline_level
            new_dim.hidden = hidden
            work_sequence = get_column_interval(start, end)[1:]
            for column_letter in work_sequence:
                if column_letter in self:
                    del self[column_letter]
            new_dim.min, new_dim.max = map(column_index_from_string, (start, end))
        elif isinstance(self.default_factory(), RowDimension):
            for el in range(start, end + 1):
                new_dim = self.worksheet.row_dimensions[el]
                new_dim.outline_level = outline_level
                new_dim.hidden = hidden


    def to_tree(self):

        def sorter(value):
            value.reindex()
            return value.min

        el = Element('cols')
        outlines = set()

        for col in sorted(self.values(), key=sorter):
            obj = col.to_tree()
            if obj is not None:
                outlines.add(col.outlineLevel)
                el.append(obj)

        if outlines:
            self.max_outline = max(outlines)

        if len(el):
            return el # must have at least one child


class SheetFormatProperties(Serialisable):

    tagname = "sheetFormatPr"

    baseColWidth = Integer(allow_none=True)
    defaultColWidth = Float(allow_none=True)
    defaultRowHeight = Float()
    customHeight = Bool(allow_none=True)
    zeroHeight = Bool(allow_none=True)
    thickTop = Bool(allow_none=True)
    thickBottom = Bool(allow_none=True)
    outlineLevelRow = Integer(allow_none=True)
    outlineLevelCol = Integer(allow_none=True)

    def __init__(self,
                 baseColWidth=8, #according to spec
                 defaultColWidth=None,
                 defaultRowHeight=15,
                 customHeight=None,
                 zeroHeight=None,
                 thickTop=None,
                 thickBottom=None,
                 outlineLevelRow=None,
                 outlineLevelCol=None,
                ):
        self.baseColWidth = baseColWidth
        self.defaultColWidth = defaultColWidth
        self.defaultRowHeight = defaultRowHeight
        self.customHeight = customHeight
        self.zeroHeight = zeroHeight
        self.thickTop = thickTop
        self.thickBottom = thickBottom
        self.outlineLevelRow = outlineLevelRow
        self.outlineLevelCol = outlineLevelCol


class SheetDimension(Serialisable):

    tagname = "dimension"

    ref = String()

    def __init__(self,
                 ref=None,
                ):
        self.ref = ref


    @property
    def boundaries(self):
        return range_boundaries(self.ref)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\mysql\mysqldb.py ===
# dialects/mysql/mysqldb.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors


"""

.. dialect:: mysql+mysqldb
    :name: mysqlclient (maintained fork of MySQL-Python)
    :dbapi: mysqldb
    :connectstring: mysql+mysqldb://<user>:<password>@<host>[:<port>]/<dbname>
    :url: https://pypi.org/project/mysqlclient/

Driver Status
-------------

The mysqlclient DBAPI is a maintained fork of the
`MySQL-Python <https://sourceforge.net/projects/mysql-python>`_ DBAPI
that is no longer maintained.  `mysqlclient`_ supports Python 2 and Python 3
and is very stable.

.. _mysqlclient: https://github.com/PyMySQL/mysqlclient-python

.. _mysqldb_unicode:

Unicode
-------

Please see :ref:`mysql_unicode` for current recommendations on unicode
handling.

.. _mysqldb_ssl:

SSL Connections
----------------

The mysqlclient and PyMySQL DBAPIs accept an additional dictionary under the
key "ssl", which may be specified using the
:paramref:`_sa.create_engine.connect_args` dictionary::

    engine = create_engine(
        "mysql+mysqldb://scott:tiger@192.168.0.134/test",
        connect_args={
            "ssl": {
                "ca": "/home/gord/client-ssl/ca.pem",
                "cert": "/home/gord/client-ssl/client-cert.pem",
                "key": "/home/gord/client-ssl/client-key.pem",
            }
        },
    )

For convenience, the following keys may also be specified inline within the URL
where they will be interpreted into the "ssl" dictionary automatically:
"ssl_ca", "ssl_cert", "ssl_key", "ssl_capath", "ssl_cipher",
"ssl_check_hostname". An example is as follows::

    connection_uri = (
        "mysql+mysqldb://scott:tiger@192.168.0.134/test"
        "?ssl_ca=/home/gord/client-ssl/ca.pem"
        "&ssl_cert=/home/gord/client-ssl/client-cert.pem"
        "&ssl_key=/home/gord/client-ssl/client-key.pem"
    )

.. seealso::

    :ref:`pymysql_ssl` in the PyMySQL dialect


Using MySQLdb with Google Cloud SQL
-----------------------------------

Google Cloud SQL now recommends use of the MySQLdb dialect.  Connect
using a URL like the following:

.. sourcecode:: text

    mysql+mysqldb://root@/<dbname>?unix_socket=/cloudsql/<projectid>:<instancename>

Server Side Cursors
-------------------

The mysqldb dialect supports server-side cursors. See :ref:`mysql_ss_cursors`.

"""

import re

from .base import MySQLCompiler
from .base import MySQLDialect
from .base import MySQLExecutionContext
from .base import MySQLIdentifierPreparer
from .base import TEXT
from ... import sql
from ... import util


class MySQLExecutionContext_mysqldb(MySQLExecutionContext):
    pass


class MySQLCompiler_mysqldb(MySQLCompiler):
    pass


class MySQLDialect_mysqldb(MySQLDialect):
    driver = "mysqldb"
    supports_statement_cache = True
    supports_unicode_statements = True
    supports_sane_rowcount = True
    supports_sane_multi_rowcount = True

    supports_native_decimal = True

    default_paramstyle = "format"
    execution_ctx_cls = MySQLExecutionContext_mysqldb
    statement_compiler = MySQLCompiler_mysqldb
    preparer = MySQLIdentifierPreparer

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._mysql_dbapi_version = (
            self._parse_dbapi_version(self.dbapi.__version__)
            if self.dbapi is not None and hasattr(self.dbapi, "__version__")
            else (0, 0, 0)
        )

    def _parse_dbapi_version(self, version):
        m = re.match(r"(\d+)\.(\d+)(?:\.(\d+))?", version)
        if m:
            return tuple(int(x) for x in m.group(1, 2, 3) if x is not None)
        else:
            return (0, 0, 0)

    @util.langhelpers.memoized_property
    def supports_server_side_cursors(self):
        try:
            cursors = __import__("MySQLdb.cursors").cursors
            self._sscursor = cursors.SSCursor
            return True
        except (ImportError, AttributeError):
            return False

    @classmethod
    def import_dbapi(cls):
        return __import__("MySQLdb")

    def on_connect(self):
        super_ = super().on_connect()

        def on_connect(conn):
            if super_ is not None:
                super_(conn)

            charset_name = conn.character_set_name()

            if charset_name is not None:
                cursor = conn.cursor()
                cursor.execute("SET NAMES %s" % charset_name)
                cursor.close()

        return on_connect

    def do_ping(self, dbapi_connection):
        dbapi_connection.ping()
        return True

    def do_executemany(self, cursor, statement, parameters, context=None):
        rowcount = cursor.executemany(statement, parameters)
        if context is not None:
            context._rowcount = rowcount

    def _check_unicode_returns(self, connection):
        # work around issue fixed in
        # https://github.com/farcepest/MySQLdb1/commit/cd44524fef63bd3fcb71947392326e9742d520e8
        # specific issue w/ the utf8mb4_bin collation and unicode returns

        collation = connection.exec_driver_sql(
            "show collation where %s = 'utf8mb4' and %s = 'utf8mb4_bin'"
            % (
                self.identifier_preparer.quote("Charset"),
                self.identifier_preparer.quote("Collation"),
            )
        ).scalar()
        has_utf8mb4_bin = self.server_version_info > (5,) and collation
        if has_utf8mb4_bin:
            additional_tests = [
                sql.collate(
                    sql.cast(
                        sql.literal_column("'test collated returns'"),
                        TEXT(charset="utf8mb4"),
                    ),
                    "utf8mb4_bin",
                )
            ]
        else:
            additional_tests = []
        return super()._check_unicode_returns(connection, additional_tests)

    def create_connect_args(self, url, _translate_args=None):
        if _translate_args is None:
            _translate_args = dict(
                database="db", username="user", password="passwd"
            )

        opts = url.translate_connect_args(**_translate_args)
        opts.update(url.query)

        util.coerce_kw_type(opts, "compress", bool)
        util.coerce_kw_type(opts, "connect_timeout", int)
        util.coerce_kw_type(opts, "read_timeout", int)
        util.coerce_kw_type(opts, "write_timeout", int)
        util.coerce_kw_type(opts, "client_flag", int)
        util.coerce_kw_type(opts, "local_infile", bool)
        # Note: using either of the below will cause all strings to be
        # returned as Unicode, both in raw SQL operations and with column
        # types like String and MSString.
        util.coerce_kw_type(opts, "use_unicode", bool)
        util.coerce_kw_type(opts, "charset", str)

        # Rich values 'cursorclass' and 'conv' are not supported via
        # query string.

        ssl = {}
        keys = [
            ("ssl_ca", str),
            ("ssl_key", str),
            ("ssl_cert", str),
            ("ssl_capath", str),
            ("ssl_cipher", str),
            ("ssl_check_hostname", bool),
        ]
        for key, kw_type in keys:
            if key in opts:
                ssl[key[4:]] = opts[key]
                util.coerce_kw_type(ssl, key[4:], kw_type)
                del opts[key]
        if ssl:
            opts["ssl"] = ssl

        # FOUND_ROWS must be set in CLIENT_FLAGS to enable
        # supports_sane_rowcount.
        client_flag = opts.get("client_flag", 0)

        client_flag_found_rows = self._found_rows_client_flag()
        if client_flag_found_rows is not None:
            client_flag |= client_flag_found_rows
            opts["client_flag"] = client_flag
        return [[], opts]

    def _found_rows_client_flag(self):
        if self.dbapi is not None:
            try:
                CLIENT_FLAGS = __import__(
                    self.dbapi.__name__ + ".constants.CLIENT"
                ).constants.CLIENT
            except (AttributeError, ImportError):
                return None
            else:
                return CLIENT_FLAGS.FOUND_ROWS
        else:
            return None

    def _extract_error_code(self, exception):
        return exception.args[0]

    def _detect_charset(self, connection):
        """Sniff out the character set in use for connection results."""

        try:
            # note: the SQL here would be
            # "SHOW VARIABLES LIKE 'character_set%%'"
            cset_name = connection.connection.character_set_name
        except AttributeError:
            util.warn(
                "No 'character_set_name' can be detected with "
                "this MySQL-Python version; "
                "please upgrade to a recent version of MySQL-Python.  "
                "Assuming latin1."
            )
            return "latin1"
        else:
            return cset_name()

    def get_isolation_level_values(self, dbapi_connection):
        return (
            "SERIALIZABLE",
            "READ UNCOMMITTED",
            "READ COMMITTED",
            "REPEATABLE READ",
            "AUTOCOMMIT",
        )

    def set_isolation_level(self, dbapi_connection, level):
        if level == "AUTOCOMMIT":
            dbapi_connection.autocommit(True)
        else:
            dbapi_connection.autocommit(False)
            super().set_isolation_level(dbapi_connection, level)


dialect = MySQLDialect_mysqldb

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\groupby\indexing.py ===
from __future__ import annotations

from collections.abc import Iterable
from typing import (
    TYPE_CHECKING,
    Literal,
    cast,
)

import numpy as np

from pandas.util._decorators import (
    cache_readonly,
    doc,
)

from pandas.core.dtypes.common import (
    is_integer,
    is_list_like,
)

if TYPE_CHECKING:
    from pandas._typing import PositionalIndexer

    from pandas import (
        DataFrame,
        Series,
    )
    from pandas.core.groupby import groupby


class GroupByIndexingMixin:
    """
    Mixin for adding ._positional_selector to GroupBy.
    """

    @cache_readonly
    def _positional_selector(self) -> GroupByPositionalSelector:
        """
        Return positional selection for each group.

        ``groupby._positional_selector[i:j]`` is similar to
        ``groupby.apply(lambda x: x.iloc[i:j])``
        but much faster and preserves the original index and order.

        ``_positional_selector[]`` is compatible with and extends :meth:`~GroupBy.head`
        and :meth:`~GroupBy.tail`. For example:

        - ``head(5)``
        - ``_positional_selector[5:-5]``
        - ``tail(5)``

        together return all the rows.

        Allowed inputs for the index are:

        - An integer valued iterable, e.g. ``range(2, 4)``.
        - A comma separated list of integers and slices, e.g. ``5``, ``2, 4``, ``2:4``.

        The output format is the same as :meth:`~GroupBy.head` and
        :meth:`~GroupBy.tail`, namely
        a subset of the ``DataFrame`` or ``Series`` with the index and order preserved.

        Returns
        -------
        Series
            The filtered subset of the original Series.
        DataFrame
            The filtered subset of the original DataFrame.

        See Also
        --------
        DataFrame.iloc : Purely integer-location based indexing for selection by
            position.
        GroupBy.head : Return first n rows of each group.
        GroupBy.tail : Return last n rows of each group.
        GroupBy.nth : Take the nth row from each group if n is an int, or a
            subset of rows, if n is a list of ints.

        Notes
        -----
        - The slice step cannot be negative.
        - If the index specification results in overlaps, the item is not duplicated.
        - If the index specification changes the order of items, then
          they are returned in their original order.
          By contrast, ``DataFrame.iloc`` can change the row order.
        - ``groupby()`` parameters such as as_index and dropna are ignored.

        The differences between ``_positional_selector[]`` and :meth:`~GroupBy.nth`
        with ``as_index=False`` are:

        - Input to ``_positional_selector`` can include
          one or more slices whereas ``nth``
          just handles an integer or a list of integers.
        - ``_positional_selector`` can  accept a slice relative to the
          last row of each group.
        - ``_positional_selector`` does not have an equivalent to the
          ``nth()`` ``dropna`` parameter.

        Examples
        --------
        >>> df = pd.DataFrame([["a", 1], ["a", 2], ["a", 3], ["b", 4], ["b", 5]],
        ...                   columns=["A", "B"])
        >>> df.groupby("A")._positional_selector[1:2]
           A  B
        1  a  2
        4  b  5

        >>> df.groupby("A")._positional_selector[1, -1]
           A  B
        1  a  2
        2  a  3
        4  b  5
        """
        if TYPE_CHECKING:
            # pylint: disable-next=used-before-assignment
            groupby_self = cast(groupby.GroupBy, self)
        else:
            groupby_self = self

        return GroupByPositionalSelector(groupby_self)

    def _make_mask_from_positional_indexer(
        self,
        arg: PositionalIndexer | tuple,
    ) -> np.ndarray:
        if is_list_like(arg):
            if all(is_integer(i) for i in cast(Iterable, arg)):
                mask = self._make_mask_from_list(cast(Iterable[int], arg))
            else:
                mask = self._make_mask_from_tuple(cast(tuple, arg))

        elif isinstance(arg, slice):
            mask = self._make_mask_from_slice(arg)
        elif is_integer(arg):
            mask = self._make_mask_from_int(cast(int, arg))
        else:
            raise TypeError(
                f"Invalid index {type(arg)}. "
                "Must be integer, list-like, slice or a tuple of "
                "integers and slices"
            )

        if isinstance(mask, bool):
            if mask:
                mask = self._ascending_count >= 0
            else:
                mask = self._ascending_count < 0

        return cast(np.ndarray, mask)

    def _make_mask_from_int(self, arg: int) -> np.ndarray:
        if arg >= 0:
            return self._ascending_count == arg
        else:
            return self._descending_count == (-arg - 1)

    def _make_mask_from_list(self, args: Iterable[int]) -> bool | np.ndarray:
        positive = [arg for arg in args if arg >= 0]
        negative = [-arg - 1 for arg in args if arg < 0]

        mask: bool | np.ndarray = False

        if positive:
            mask |= np.isin(self._ascending_count, positive)

        if negative:
            mask |= np.isin(self._descending_count, negative)

        return mask

    def _make_mask_from_tuple(self, args: tuple) -> bool | np.ndarray:
        mask: bool | np.ndarray = False

        for arg in args:
            if is_integer(arg):
                mask |= self._make_mask_from_int(cast(int, arg))
            elif isinstance(arg, slice):
                mask |= self._make_mask_from_slice(arg)
            else:
                raise ValueError(
                    f"Invalid argument {type(arg)}. Should be int or slice."
                )

        return mask

    def _make_mask_from_slice(self, arg: slice) -> bool | np.ndarray:
        start = arg.start
        stop = arg.stop
        step = arg.step

        if step is not None and step < 0:
            raise ValueError(f"Invalid step {step}. Must be non-negative")

        mask: bool | np.ndarray = True

        if step is None:
            step = 1

        if start is None:
            if step > 1:
                mask &= self._ascending_count % step == 0

        elif start >= 0:
            mask &= self._ascending_count >= start

            if step > 1:
                mask &= (self._ascending_count - start) % step == 0

        else:
            mask &= self._descending_count < -start

            offset_array = self._descending_count + start + 1
            limit_array = (
                self._ascending_count + self._descending_count + (start + 1)
            ) < 0
            offset_array = np.where(limit_array, self._ascending_count, offset_array)

            mask &= offset_array % step == 0

        if stop is not None:
            if stop >= 0:
                mask &= self._ascending_count < stop
            else:
                mask &= self._descending_count >= -stop

        return mask

    @cache_readonly
    def _ascending_count(self) -> np.ndarray:
        if TYPE_CHECKING:
            groupby_self = cast(groupby.GroupBy, self)
        else:
            groupby_self = self

        return groupby_self._cumcount_array()

    @cache_readonly
    def _descending_count(self) -> np.ndarray:
        if TYPE_CHECKING:
            groupby_self = cast(groupby.GroupBy, self)
        else:
            groupby_self = self

        return groupby_self._cumcount_array(ascending=False)


@doc(GroupByIndexingMixin._positional_selector)
class GroupByPositionalSelector:
    def __init__(self, groupby_object: groupby.GroupBy) -> None:
        self.groupby_object = groupby_object

    def __getitem__(self, arg: PositionalIndexer | tuple) -> DataFrame | Series:
        """
        Select by positional index per group.

        Implements GroupBy._positional_selector

        Parameters
        ----------
        arg : PositionalIndexer | tuple
            Allowed values are:
            - int
            - int valued iterable such as list or range
            - slice with step either None or positive
            - tuple of integers and slices

        Returns
        -------
        Series
            The filtered subset of the original groupby Series.
        DataFrame
            The filtered subset of the original groupby DataFrame.

        See Also
        --------
        DataFrame.iloc : Integer-location based indexing for selection by position.
        GroupBy.head : Return first n rows of each group.
        GroupBy.tail : Return last n rows of each group.
        GroupBy._positional_selector : Return positional selection for each group.
        GroupBy.nth : Take the nth row from each group if n is an int, or a
            subset of rows, if n is a list of ints.
        """
        mask = self.groupby_object._make_mask_from_positional_indexer(arg)
        return self.groupby_object._mask_selected_obj(mask)


class GroupByNthSelector:
    """
    Dynamically substituted for GroupBy.nth to enable both call and index
    """

    def __init__(self, groupby_object: groupby.GroupBy) -> None:
        self.groupby_object = groupby_object

    def __call__(
        self,
        n: PositionalIndexer | tuple,
        dropna: Literal["any", "all", None] = None,
    ) -> DataFrame | Series:
        return self.groupby_object._nth(n, dropna)

    def __getitem__(self, n: PositionalIndexer | tuple) -> DataFrame | Series:
        return self.groupby_object._nth(n)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\core\traversal.py ===
from __future__ import annotations

from typing import Iterator

from .basic import Basic
from .sorting import ordered
from .sympify import sympify
from sympy.utilities.iterables import iterable



def iterargs(expr):
    """Yield the args of a Basic object in a breadth-first traversal.
    Depth-traversal stops if `arg.args` is either empty or is not
    an iterable.

    Examples
    ========

    >>> from sympy import Integral, Function
    >>> from sympy.abc import x
    >>> f = Function('f')
    >>> from sympy.core.traversal import iterargs
    >>> list(iterargs(Integral(f(x), (f(x), 1))))
    [Integral(f(x), (f(x), 1)), f(x), (f(x), 1), x, f(x), 1, x]

    See Also
    ========
    iterfreeargs, preorder_traversal
    """
    args = [expr]
    for i in args:
        yield i
        args.extend(i.args)


def iterfreeargs(expr, _first=True):
    """Yield the args of a Basic object in a breadth-first traversal.
    Depth-traversal stops if `arg.args` is either empty or is not
    an iterable. The bound objects of an expression will be returned
    as canonical variables.

    Examples
    ========

    >>> from sympy import Integral, Function
    >>> from sympy.abc import x
    >>> f = Function('f')
    >>> from sympy.core.traversal import iterfreeargs
    >>> list(iterfreeargs(Integral(f(x), (f(x), 1))))
    [Integral(f(x), (f(x), 1)), 1]

    See Also
    ========
    iterargs, preorder_traversal
    """
    args = [expr]
    for i in args:
        yield i
        if _first and hasattr(i, 'bound_symbols'):
            void = i.canonical_variables.values()
            for i in iterfreeargs(i.as_dummy(), _first=False):
                if not i.has(*void):
                    yield i
        args.extend(i.args)


class preorder_traversal:
    """
    Do a pre-order traversal of a tree.

    This iterator recursively yields nodes that it has visited in a pre-order
    fashion. That is, it yields the current node then descends through the
    tree breadth-first to yield all of a node's children's pre-order
    traversal.


    For an expression, the order of the traversal depends on the order of
    .args, which in many cases can be arbitrary.

    Parameters
    ==========
    node : SymPy expression
        The expression to traverse.
    keys : (default None) sort key(s)
        The key(s) used to sort args of Basic objects. When None, args of Basic
        objects are processed in arbitrary order. If key is defined, it will
        be passed along to ordered() as the only key(s) to use to sort the
        arguments; if ``key`` is simply True then the default keys of ordered
        will be used.

    Yields
    ======
    subtree : SymPy expression
        All of the subtrees in the tree.

    Examples
    ========

    >>> from sympy import preorder_traversal, symbols
    >>> x, y, z = symbols('x y z')

    The nodes are returned in the order that they are encountered unless key
    is given; simply passing key=True will guarantee that the traversal is
    unique.

    >>> list(preorder_traversal((x + y)*z, keys=None)) # doctest: +SKIP
    [z*(x + y), z, x + y, y, x]
    >>> list(preorder_traversal((x + y)*z, keys=True))
    [z*(x + y), z, x + y, x, y]

    """
    def __init__(self, node, keys=None):
        self._skip_flag = False
        self._pt = self._preorder_traversal(node, keys)

    def _preorder_traversal(self, node, keys):
        yield node
        if self._skip_flag:
            self._skip_flag = False
            return
        if isinstance(node, Basic):
            if not keys and hasattr(node, '_argset'):
                # LatticeOp keeps args as a set. We should use this if we
                # don't care about the order, to prevent unnecessary sorting.
                args = node._argset
            else:
                args = node.args
            if keys:
                if keys != True:
                    args = ordered(args, keys, default=False)
                else:
                    args = ordered(args)
            for arg in args:
                yield from self._preorder_traversal(arg, keys)
        elif iterable(node):
            for item in node:
                yield from self._preorder_traversal(item, keys)

    def skip(self):
        """
        Skip yielding current node's (last yielded node's) subtrees.

        Examples
        ========

        >>> from sympy import preorder_traversal, symbols
        >>> x, y, z = symbols('x y z')
        >>> pt = preorder_traversal((x + y*z)*z)
        >>> for i in pt:
        ...     print(i)
        ...     if i == x + y*z:
        ...             pt.skip()
        z*(x + y*z)
        z
        x + y*z
        """
        self._skip_flag = True

    def __next__(self):
        return next(self._pt)

    def __iter__(self) -> Iterator[Basic]:
        return self


def use(expr, func, level=0, args=(), kwargs={}):
    """
    Use ``func`` to transform ``expr`` at the given level.

    Examples
    ========

    >>> from sympy import use, expand
    >>> from sympy.abc import x, y

    >>> f = (x + y)**2*x + 1

    >>> use(f, expand, level=2)
    x*(x**2 + 2*x*y + y**2) + 1
    >>> expand(f)
    x**3 + 2*x**2*y + x*y**2 + 1

    """
    def _use(expr, level):
        if not level:
            return func(expr, *args, **kwargs)
        else:
            if expr.is_Atom:
                return expr
            else:
                level -= 1
                _args = [_use(arg, level) for arg in expr.args]
                return expr.__class__(*_args)

    return _use(sympify(expr), level)


def walk(e, *target):
    """Iterate through the args that are the given types (target) and
    return a list of the args that were traversed; arguments
    that are not of the specified types are not traversed.

    Examples
    ========

    >>> from sympy.core.traversal import walk
    >>> from sympy import Min, Max
    >>> from sympy.abc import x, y, z
    >>> list(walk(Min(x, Max(y, Min(1, z))), Min))
    [Min(x, Max(y, Min(1, z)))]
    >>> list(walk(Min(x, Max(y, Min(1, z))), Min, Max))
    [Min(x, Max(y, Min(1, z))), Max(y, Min(1, z)), Min(1, z)]

    See Also
    ========

    bottom_up
    """
    if isinstance(e, target):
        yield e
        for i in e.args:
            yield from walk(i, *target)


def bottom_up(rv, F, atoms=False, nonbasic=False):
    """Apply ``F`` to all expressions in an expression tree from the
    bottom up. If ``atoms`` is True, apply ``F`` even if there are no args;
    if ``nonbasic`` is True, try to apply ``F`` to non-Basic objects.
    """
    args = getattr(rv, 'args', None)
    if args is not None:
        if args:
            args = tuple([bottom_up(a, F, atoms, nonbasic) for a in args])
            if args != rv.args:
                rv = rv.func(*args)
            rv = F(rv)
        elif atoms:
            rv = F(rv)
    else:
        if nonbasic:
            try:
                rv = F(rv)
            except TypeError:
                pass

    return rv


def postorder_traversal(node, keys=None):
    """
    Do a postorder traversal of a tree.

    This generator recursively yields nodes that it has visited in a postorder
    fashion. That is, it descends through the tree depth-first to yield all of
    a node's children's postorder traversal before yielding the node itself.

    Parameters
    ==========

    node : SymPy expression
        The expression to traverse.
    keys : (default None) sort key(s)
        The key(s) used to sort args of Basic objects. When None, args of Basic
        objects are processed in arbitrary order. If key is defined, it will
        be passed along to ordered() as the only key(s) to use to sort the
        arguments; if ``key`` is simply True then the default keys of
        ``ordered`` will be used (node count and default_sort_key).

    Yields
    ======
    subtree : SymPy expression
        All of the subtrees in the tree.

    Examples
    ========

    >>> from sympy import postorder_traversal
    >>> from sympy.abc import w, x, y, z

    The nodes are returned in the order that they are encountered unless key
    is given; simply passing key=True will guarantee that the traversal is
    unique.

    >>> list(postorder_traversal(w + (x + y)*z)) # doctest: +SKIP
    [z, y, x, x + y, z*(x + y), w, w + z*(x + y)]
    >>> list(postorder_traversal(w + (x + y)*z, keys=True))
    [w, z, x, y, x + y, z*(x + y), w + z*(x + y)]


    """
    if isinstance(node, Basic):
        args = node.args
        if keys:
            if keys != True:
                args = ordered(args, keys, default=False)
            else:
                args = ordered(args)
        for arg in args:
            yield from postorder_traversal(arg, keys)
    elif iterable(node):
        for item in node:
            yield from postorder_traversal(item, keys)
    yield node

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v135\memory.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: Memory (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

class PressureLevel(enum.Enum):
    '''
    Memory pressure level.
    '''
    MODERATE = "moderate"
    CRITICAL = "critical"

    def to_json(self):
        return self.value

    @classmethod
    def from_json(cls, json):
        return cls(json)


@dataclass
class SamplingProfileNode:
    '''
    Heap profile sample.
    '''
    #: Size of the sampled allocation.
    size: float

    #: Total bytes attributed to this sample.
    total: float

    #: Execution stack at the point of allocation.
    stack: typing.List[str]

    def to_json(self):
        json = dict()
        json['size'] = self.size
        json['total'] = self.total
        json['stack'] = [i for i in self.stack]
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            size=float(json['size']),
            total=float(json['total']),
            stack=[str(i) for i in json['stack']],
        )


@dataclass
class SamplingProfile:
    '''
    Array of heap profile samples.
    '''
    samples: typing.List[SamplingProfileNode]

    modules: typing.List[Module]

    def to_json(self):
        json = dict()
        json['samples'] = [i.to_json() for i in self.samples]
        json['modules'] = [i.to_json() for i in self.modules]
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            samples=[SamplingProfileNode.from_json(i) for i in json['samples']],
            modules=[Module.from_json(i) for i in json['modules']],
        )


@dataclass
class Module:
    '''
    Executable module information
    '''
    #: Name of the module.
    name: str

    #: UUID of the module.
    uuid: str

    #: Base address where the module is loaded into memory. Encoded as a decimal
    #: or hexadecimal (0x prefixed) string.
    base_address: str

    #: Size of the module in bytes.
    size: float

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['uuid'] = self.uuid
        json['baseAddress'] = self.base_address
        json['size'] = self.size
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            uuid=str(json['uuid']),
            base_address=str(json['baseAddress']),
            size=float(json['size']),
        )


@dataclass
class DOMCounter:
    '''
    DOM object counter data.
    '''
    #: Object name. Note: object names should be presumed volatile and clients should not expect
    #: the returned names to be consistent across runs.
    name: str

    #: Object count.
    count: int

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['count'] = self.count
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            count=int(json['count']),
        )


def get_dom_counters() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.Tuple[int, int, int]]:
    '''
    Retruns current DOM object counters.

    :returns: A tuple with the following items:

        0. **documents** -
        1. **nodes** -
        2. **jsEventListeners** -
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getDOMCounters',
    }
    json = yield cmd_dict
    return (
        int(json['documents']),
        int(json['nodes']),
        int(json['jsEventListeners'])
    )


def get_dom_counters_for_leak_detection() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[DOMCounter]]:
    '''
    Retruns DOM object counters after preparing renderer for leak detection.

    :returns: DOM object counters.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getDOMCountersForLeakDetection',
    }
    json = yield cmd_dict
    return [DOMCounter.from_json(i) for i in json['counters']]


def prepare_for_leak_detection() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Prepares for leak detection by terminating workers, stopping spellcheckers,
    dropping non-essential internal caches, running garbage collections, etc.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.prepareForLeakDetection',
    }
    json = yield cmd_dict


def forcibly_purge_java_script_memory() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Simulate OomIntervention by purging V8 memory.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.forciblyPurgeJavaScriptMemory',
    }
    json = yield cmd_dict


def set_pressure_notifications_suppressed(
        suppressed: bool
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enable/disable suppressing memory pressure notifications in all processes.

    :param suppressed: If true, memory pressure notifications will be suppressed.
    '''
    params: T_JSON_DICT = dict()
    params['suppressed'] = suppressed
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.setPressureNotificationsSuppressed',
        'params': params,
    }
    json = yield cmd_dict


def simulate_pressure_notification(
        level: PressureLevel
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Simulate a memory pressure notification in all processes.

    :param level: Memory pressure level of the notification.
    '''
    params: T_JSON_DICT = dict()
    params['level'] = level.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.simulatePressureNotification',
        'params': params,
    }
    json = yield cmd_dict


def start_sampling(
        sampling_interval: typing.Optional[int] = None,
        suppress_randomness: typing.Optional[bool] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Start collecting native memory profile.

    :param sampling_interval: *(Optional)* Average number of bytes between samples.
    :param suppress_randomness: *(Optional)* Do not randomize intervals between samples.
    '''
    params: T_JSON_DICT = dict()
    if sampling_interval is not None:
        params['samplingInterval'] = sampling_interval
    if suppress_randomness is not None:
        params['suppressRandomness'] = suppress_randomness
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.startSampling',
        'params': params,
    }
    json = yield cmd_dict


def stop_sampling() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stop collecting native memory profile.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.stopSampling',
    }
    json = yield cmd_dict


def get_all_time_sampling_profile() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SamplingProfile]:
    '''
    Retrieve native memory allocations profile
    collected since renderer process startup.

    :returns:
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getAllTimeSamplingProfile',
    }
    json = yield cmd_dict
    return SamplingProfile.from_json(json['profile'])


def get_browser_sampling_profile() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SamplingProfile]:
    '''
    Retrieve native memory allocations profile
    collected since browser process startup.

    :returns:
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getBrowserSamplingProfile',
    }
    json = yield cmd_dict
    return SamplingProfile.from_json(json['profile'])


def get_sampling_profile() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SamplingProfile]:
    '''
    Retrieve native memory allocations profile collected since last
    ``startSampling`` call.

    :returns:
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getSamplingProfile',
    }
    json = yield cmd_dict
    return SamplingProfile.from_json(json['profile'])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v136\memory.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: Memory (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

class PressureLevel(enum.Enum):
    '''
    Memory pressure level.
    '''
    MODERATE = "moderate"
    CRITICAL = "critical"

    def to_json(self):
        return self.value

    @classmethod
    def from_json(cls, json):
        return cls(json)


@dataclass
class SamplingProfileNode:
    '''
    Heap profile sample.
    '''
    #: Size of the sampled allocation.
    size: float

    #: Total bytes attributed to this sample.
    total: float

    #: Execution stack at the point of allocation.
    stack: typing.List[str]

    def to_json(self):
        json = dict()
        json['size'] = self.size
        json['total'] = self.total
        json['stack'] = [i for i in self.stack]
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            size=float(json['size']),
            total=float(json['total']),
            stack=[str(i) for i in json['stack']],
        )


@dataclass
class SamplingProfile:
    '''
    Array of heap profile samples.
    '''
    samples: typing.List[SamplingProfileNode]

    modules: typing.List[Module]

    def to_json(self):
        json = dict()
        json['samples'] = [i.to_json() for i in self.samples]
        json['modules'] = [i.to_json() for i in self.modules]
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            samples=[SamplingProfileNode.from_json(i) for i in json['samples']],
            modules=[Module.from_json(i) for i in json['modules']],
        )


@dataclass
class Module:
    '''
    Executable module information
    '''
    #: Name of the module.
    name: str

    #: UUID of the module.
    uuid: str

    #: Base address where the module is loaded into memory. Encoded as a decimal
    #: or hexadecimal (0x prefixed) string.
    base_address: str

    #: Size of the module in bytes.
    size: float

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['uuid'] = self.uuid
        json['baseAddress'] = self.base_address
        json['size'] = self.size
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            uuid=str(json['uuid']),
            base_address=str(json['baseAddress']),
            size=float(json['size']),
        )


@dataclass
class DOMCounter:
    '''
    DOM object counter data.
    '''
    #: Object name. Note: object names should be presumed volatile and clients should not expect
    #: the returned names to be consistent across runs.
    name: str

    #: Object count.
    count: int

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['count'] = self.count
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            count=int(json['count']),
        )


def get_dom_counters() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.Tuple[int, int, int]]:
    '''
    Retruns current DOM object counters.

    :returns: A tuple with the following items:

        0. **documents** -
        1. **nodes** -
        2. **jsEventListeners** -
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getDOMCounters',
    }
    json = yield cmd_dict
    return (
        int(json['documents']),
        int(json['nodes']),
        int(json['jsEventListeners'])
    )


def get_dom_counters_for_leak_detection() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[DOMCounter]]:
    '''
    Retruns DOM object counters after preparing renderer for leak detection.

    :returns: DOM object counters.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getDOMCountersForLeakDetection',
    }
    json = yield cmd_dict
    return [DOMCounter.from_json(i) for i in json['counters']]


def prepare_for_leak_detection() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Prepares for leak detection by terminating workers, stopping spellcheckers,
    dropping non-essential internal caches, running garbage collections, etc.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.prepareForLeakDetection',
    }
    json = yield cmd_dict


def forcibly_purge_java_script_memory() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Simulate OomIntervention by purging V8 memory.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.forciblyPurgeJavaScriptMemory',
    }
    json = yield cmd_dict


def set_pressure_notifications_suppressed(
        suppressed: bool
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enable/disable suppressing memory pressure notifications in all processes.

    :param suppressed: If true, memory pressure notifications will be suppressed.
    '''
    params: T_JSON_DICT = dict()
    params['suppressed'] = suppressed
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.setPressureNotificationsSuppressed',
        'params': params,
    }
    json = yield cmd_dict


def simulate_pressure_notification(
        level: PressureLevel
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Simulate a memory pressure notification in all processes.

    :param level: Memory pressure level of the notification.
    '''
    params: T_JSON_DICT = dict()
    params['level'] = level.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.simulatePressureNotification',
        'params': params,
    }
    json = yield cmd_dict


def start_sampling(
        sampling_interval: typing.Optional[int] = None,
        suppress_randomness: typing.Optional[bool] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Start collecting native memory profile.

    :param sampling_interval: *(Optional)* Average number of bytes between samples.
    :param suppress_randomness: *(Optional)* Do not randomize intervals between samples.
    '''
    params: T_JSON_DICT = dict()
    if sampling_interval is not None:
        params['samplingInterval'] = sampling_interval
    if suppress_randomness is not None:
        params['suppressRandomness'] = suppress_randomness
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.startSampling',
        'params': params,
    }
    json = yield cmd_dict


def stop_sampling() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stop collecting native memory profile.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.stopSampling',
    }
    json = yield cmd_dict


def get_all_time_sampling_profile() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SamplingProfile]:
    '''
    Retrieve native memory allocations profile
    collected since renderer process startup.

    :returns:
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getAllTimeSamplingProfile',
    }
    json = yield cmd_dict
    return SamplingProfile.from_json(json['profile'])


def get_browser_sampling_profile() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SamplingProfile]:
    '''
    Retrieve native memory allocations profile
    collected since browser process startup.

    :returns:
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getBrowserSamplingProfile',
    }
    json = yield cmd_dict
    return SamplingProfile.from_json(json['profile'])


def get_sampling_profile() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SamplingProfile]:
    '''
    Retrieve native memory allocations profile collected since last
    ``startSampling`` call.

    :returns:
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getSamplingProfile',
    }
    json = yield cmd_dict
    return SamplingProfile.from_json(json['profile'])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v137\memory.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: Memory (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

class PressureLevel(enum.Enum):
    '''
    Memory pressure level.
    '''
    MODERATE = "moderate"
    CRITICAL = "critical"

    def to_json(self):
        return self.value

    @classmethod
    def from_json(cls, json):
        return cls(json)


@dataclass
class SamplingProfileNode:
    '''
    Heap profile sample.
    '''
    #: Size of the sampled allocation.
    size: float

    #: Total bytes attributed to this sample.
    total: float

    #: Execution stack at the point of allocation.
    stack: typing.List[str]

    def to_json(self):
        json = dict()
        json['size'] = self.size
        json['total'] = self.total
        json['stack'] = [i for i in self.stack]
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            size=float(json['size']),
            total=float(json['total']),
            stack=[str(i) for i in json['stack']],
        )


@dataclass
class SamplingProfile:
    '''
    Array of heap profile samples.
    '''
    samples: typing.List[SamplingProfileNode]

    modules: typing.List[Module]

    def to_json(self):
        json = dict()
        json['samples'] = [i.to_json() for i in self.samples]
        json['modules'] = [i.to_json() for i in self.modules]
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            samples=[SamplingProfileNode.from_json(i) for i in json['samples']],
            modules=[Module.from_json(i) for i in json['modules']],
        )


@dataclass
class Module:
    '''
    Executable module information
    '''
    #: Name of the module.
    name: str

    #: UUID of the module.
    uuid: str

    #: Base address where the module is loaded into memory. Encoded as a decimal
    #: or hexadecimal (0x prefixed) string.
    base_address: str

    #: Size of the module in bytes.
    size: float

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['uuid'] = self.uuid
        json['baseAddress'] = self.base_address
        json['size'] = self.size
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            uuid=str(json['uuid']),
            base_address=str(json['baseAddress']),
            size=float(json['size']),
        )


@dataclass
class DOMCounter:
    '''
    DOM object counter data.
    '''
    #: Object name. Note: object names should be presumed volatile and clients should not expect
    #: the returned names to be consistent across runs.
    name: str

    #: Object count.
    count: int

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['count'] = self.count
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            count=int(json['count']),
        )


def get_dom_counters() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.Tuple[int, int, int]]:
    '''
    Retruns current DOM object counters.

    :returns: A tuple with the following items:

        0. **documents** -
        1. **nodes** -
        2. **jsEventListeners** -
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getDOMCounters',
    }
    json = yield cmd_dict
    return (
        int(json['documents']),
        int(json['nodes']),
        int(json['jsEventListeners'])
    )


def get_dom_counters_for_leak_detection() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[DOMCounter]]:
    '''
    Retruns DOM object counters after preparing renderer for leak detection.

    :returns: DOM object counters.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getDOMCountersForLeakDetection',
    }
    json = yield cmd_dict
    return [DOMCounter.from_json(i) for i in json['counters']]


def prepare_for_leak_detection() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Prepares for leak detection by terminating workers, stopping spellcheckers,
    dropping non-essential internal caches, running garbage collections, etc.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.prepareForLeakDetection',
    }
    json = yield cmd_dict


def forcibly_purge_java_script_memory() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Simulate OomIntervention by purging V8 memory.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.forciblyPurgeJavaScriptMemory',
    }
    json = yield cmd_dict


def set_pressure_notifications_suppressed(
        suppressed: bool
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enable/disable suppressing memory pressure notifications in all processes.

    :param suppressed: If true, memory pressure notifications will be suppressed.
    '''
    params: T_JSON_DICT = dict()
    params['suppressed'] = suppressed
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.setPressureNotificationsSuppressed',
        'params': params,
    }
    json = yield cmd_dict


def simulate_pressure_notification(
        level: PressureLevel
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Simulate a memory pressure notification in all processes.

    :param level: Memory pressure level of the notification.
    '''
    params: T_JSON_DICT = dict()
    params['level'] = level.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.simulatePressureNotification',
        'params': params,
    }
    json = yield cmd_dict


def start_sampling(
        sampling_interval: typing.Optional[int] = None,
        suppress_randomness: typing.Optional[bool] = None
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Start collecting native memory profile.

    :param sampling_interval: *(Optional)* Average number of bytes between samples.
    :param suppress_randomness: *(Optional)* Do not randomize intervals between samples.
    '''
    params: T_JSON_DICT = dict()
    if sampling_interval is not None:
        params['samplingInterval'] = sampling_interval
    if suppress_randomness is not None:
        params['suppressRandomness'] = suppress_randomness
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.startSampling',
        'params': params,
    }
    json = yield cmd_dict


def stop_sampling() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stop collecting native memory profile.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.stopSampling',
    }
    json = yield cmd_dict


def get_all_time_sampling_profile() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SamplingProfile]:
    '''
    Retrieve native memory allocations profile
    collected since renderer process startup.

    :returns:
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getAllTimeSamplingProfile',
    }
    json = yield cmd_dict
    return SamplingProfile.from_json(json['profile'])


def get_browser_sampling_profile() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SamplingProfile]:
    '''
    Retrieve native memory allocations profile
    collected since browser process startup.

    :returns:
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getBrowserSamplingProfile',
    }
    json = yield cmd_dict
    return SamplingProfile.from_json(json['profile'])


def get_sampling_profile() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,SamplingProfile]:
    '''
    Retrieve native memory allocations profile collected since last
    ``startSampling`` call.

    :returns:
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Memory.getSamplingProfile',
    }
    json = yield cmd_dict
    return SamplingProfile.from_json(json['profile'])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\ext\mypy\plugin.py ===
# ext/mypy/plugin.py
# Copyright (C) 2021-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

"""
Mypy plugin for SQLAlchemy ORM.

"""
from __future__ import annotations

from typing import Callable
from typing import List
from typing import Optional
from typing import Tuple
from typing import Type as TypingType
from typing import Union

from mypy import nodes
from mypy.mro import calculate_mro
from mypy.mro import MroError
from mypy.nodes import Block
from mypy.nodes import ClassDef
from mypy.nodes import GDEF
from mypy.nodes import MypyFile
from mypy.nodes import NameExpr
from mypy.nodes import SymbolTable
from mypy.nodes import SymbolTableNode
from mypy.nodes import TypeInfo
from mypy.plugin import AttributeContext
from mypy.plugin import ClassDefContext
from mypy.plugin import DynamicClassDefContext
from mypy.plugin import Plugin
from mypy.plugin import SemanticAnalyzerPluginInterface
from mypy.types import get_proper_type
from mypy.types import Instance
from mypy.types import Type

from . import decl_class
from . import names
from . import util

try:
    __import__("sqlalchemy-stubs")
except ImportError:
    pass
else:
    raise ImportError(
        "The SQLAlchemy mypy plugin in SQLAlchemy "
        "2.0 does not work with sqlalchemy-stubs or "
        "sqlalchemy2-stubs installed, as well as with any other third party "
        "SQLAlchemy stubs.  Please uninstall all SQLAlchemy stubs "
        "packages."
    )


class SQLAlchemyPlugin(Plugin):
    def get_dynamic_class_hook(
        self, fullname: str
    ) -> Optional[Callable[[DynamicClassDefContext], None]]:
        if names.type_id_for_fullname(fullname) is names.DECLARATIVE_BASE:
            return _dynamic_class_hook
        return None

    def get_customize_class_mro_hook(
        self, fullname: str
    ) -> Optional[Callable[[ClassDefContext], None]]:
        return _fill_in_decorators

    def get_class_decorator_hook(
        self, fullname: str
    ) -> Optional[Callable[[ClassDefContext], None]]:
        sym = self.lookup_fully_qualified(fullname)

        if sym is not None and sym.node is not None:
            type_id = names.type_id_for_named_node(sym.node)
            if type_id is names.MAPPED_DECORATOR:
                return _cls_decorator_hook
            elif type_id in (
                names.AS_DECLARATIVE,
                names.AS_DECLARATIVE_BASE,
            ):
                return _base_cls_decorator_hook
            elif type_id is names.DECLARATIVE_MIXIN:
                return _declarative_mixin_hook

        return None

    def get_metaclass_hook(
        self, fullname: str
    ) -> Optional[Callable[[ClassDefContext], None]]:
        if names.type_id_for_fullname(fullname) is names.DECLARATIVE_META:
            # Set any classes that explicitly have metaclass=DeclarativeMeta
            # as declarative so the check in `get_base_class_hook()` works
            return _metaclass_cls_hook

        return None

    def get_base_class_hook(
        self, fullname: str
    ) -> Optional[Callable[[ClassDefContext], None]]:
        sym = self.lookup_fully_qualified(fullname)

        if (
            sym
            and isinstance(sym.node, TypeInfo)
            and util.has_declarative_base(sym.node)
        ):
            return _base_cls_hook

        return None

    def get_attribute_hook(
        self, fullname: str
    ) -> Optional[Callable[[AttributeContext], Type]]:
        if fullname.startswith(
            "sqlalchemy.orm.attributes.QueryableAttribute."
        ):
            return _queryable_getattr_hook

        return None

    def get_additional_deps(
        self, file: MypyFile
    ) -> List[Tuple[int, str, int]]:
        return [
            #
            (10, "sqlalchemy.orm", -1),
            (10, "sqlalchemy.orm.attributes", -1),
            (10, "sqlalchemy.orm.decl_api", -1),
        ]


def plugin(version: str) -> TypingType[SQLAlchemyPlugin]:
    return SQLAlchemyPlugin


def _dynamic_class_hook(ctx: DynamicClassDefContext) -> None:
    """Generate a declarative Base class when the declarative_base() function
    is encountered."""

    _add_globals(ctx)

    cls = ClassDef(ctx.name, Block([]))
    cls.fullname = ctx.api.qualified_name(ctx.name)

    info = TypeInfo(SymbolTable(), cls, ctx.api.cur_mod_id)
    cls.info = info
    _set_declarative_metaclass(ctx.api, cls)

    cls_arg = util.get_callexpr_kwarg(ctx.call, "cls", expr_types=(NameExpr,))
    if cls_arg is not None and isinstance(cls_arg.node, TypeInfo):
        util.set_is_base(cls_arg.node)
        decl_class.scan_declarative_assignments_and_apply_types(
            cls_arg.node.defn, ctx.api, is_mixin_scan=True
        )
        info.bases = [Instance(cls_arg.node, [])]
    else:
        obj = ctx.api.named_type(names.NAMED_TYPE_BUILTINS_OBJECT)

        info.bases = [obj]

    try:
        calculate_mro(info)
    except MroError:
        util.fail(
            ctx.api, "Not able to calculate MRO for declarative base", ctx.call
        )
        obj = ctx.api.named_type(names.NAMED_TYPE_BUILTINS_OBJECT)
        info.bases = [obj]
        info.fallback_to_any = True

    ctx.api.add_symbol_table_node(ctx.name, SymbolTableNode(GDEF, info))
    util.set_is_base(info)


def _fill_in_decorators(ctx: ClassDefContext) -> None:
    for decorator in ctx.cls.decorators:
        # set the ".fullname" attribute of a class decorator
        # that is a MemberExpr.   This causes the logic in
        # semanal.py->apply_class_plugin_hooks to invoke the
        # get_class_decorator_hook for our "registry.map_class()"
        # and "registry.as_declarative_base()" methods.
        # this seems like a bug in mypy that these decorators are otherwise
        # skipped.

        if (
            isinstance(decorator, nodes.CallExpr)
            and isinstance(decorator.callee, nodes.MemberExpr)
            and decorator.callee.name == "as_declarative_base"
        ):
            target = decorator.callee
        elif (
            isinstance(decorator, nodes.MemberExpr)
            and decorator.name == "mapped"
        ):
            target = decorator
        else:
            continue

        if isinstance(target.expr, NameExpr):
            sym = ctx.api.lookup_qualified(
                target.expr.name, target, suppress_errors=True
            )
        else:
            continue

        if sym and sym.node:
            sym_type = get_proper_type(sym.type)
            if isinstance(sym_type, Instance):
                target.fullname = f"{sym_type.type.fullname}.{target.name}"
            else:
                # if the registry is in the same file as where the
                # decorator is used, it might not have semantic
                # symbols applied and we can't get a fully qualified
                # name or an inferred type, so we are actually going to
                # flag an error in this case that they need to annotate
                # it.  The "registry" is declared just
                # once (or few times), so they have to just not use
                # type inference for its assignment in this one case.
                util.fail(
                    ctx.api,
                    "Class decorator called %s(), but we can't "
                    "tell if it's from an ORM registry.  Please "
                    "annotate the registry assignment, e.g. "
                    "my_registry: registry = registry()" % target.name,
                    sym.node,
                )


def _cls_decorator_hook(ctx: ClassDefContext) -> None:
    _add_globals(ctx)
    assert isinstance(ctx.reason, nodes.MemberExpr)
    expr = ctx.reason.expr

    assert isinstance(expr, nodes.RefExpr) and isinstance(expr.node, nodes.Var)

    node_type = get_proper_type(expr.node.type)

    assert (
        isinstance(node_type, Instance)
        and names.type_id_for_named_node(node_type.type) is names.REGISTRY
    )

    decl_class.scan_declarative_assignments_and_apply_types(ctx.cls, ctx.api)


def _base_cls_decorator_hook(ctx: ClassDefContext) -> None:
    _add_globals(ctx)

    cls = ctx.cls

    _set_declarative_metaclass(ctx.api, cls)

    util.set_is_base(ctx.cls.info)
    decl_class.scan_declarative_assignments_and_apply_types(
        cls, ctx.api, is_mixin_scan=True
    )


def _declarative_mixin_hook(ctx: ClassDefContext) -> None:
    _add_globals(ctx)
    util.set_is_base(ctx.cls.info)
    decl_class.scan_declarative_assignments_and_apply_types(
        ctx.cls, ctx.api, is_mixin_scan=True
    )


def _metaclass_cls_hook(ctx: ClassDefContext) -> None:
    util.set_is_base(ctx.cls.info)


def _base_cls_hook(ctx: ClassDefContext) -> None:
    _add_globals(ctx)
    decl_class.scan_declarative_assignments_and_apply_types(ctx.cls, ctx.api)


def _queryable_getattr_hook(ctx: AttributeContext) -> Type:
    # how do I....tell it it has no attribute of a certain name?
    # can't find any Type that seems to match that
    return ctx.default_attr_type


def _add_globals(ctx: Union[ClassDefContext, DynamicClassDefContext]) -> None:
    """Add __sa_DeclarativeMeta and __sa_Mapped symbol to the global space
    for all class defs

    """

    util.add_global(ctx, "sqlalchemy.orm", "Mapped", "__sa_Mapped")


def _set_declarative_metaclass(
    api: SemanticAnalyzerPluginInterface, target_cls: ClassDef
) -> None:
    info = target_cls.info
    sym = api.lookup_fully_qualified_or_none(
        "sqlalchemy.orm.decl_api.DeclarativeMeta"
    )
    assert sym is not None and isinstance(sym.node, TypeInfo)
    info.declared_metaclass = info.metaclass_type = Instance(sym.node, [])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\util\compat.py ===
# util/compat.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: allow-untyped-defs, allow-untyped-calls

"""Handle Python version/platform incompatibilities."""

from __future__ import annotations

import base64
import dataclasses
import hashlib
import inspect
import operator
import platform
import sys
import typing
from typing import Any
from typing import Callable
from typing import Dict
from typing import Iterable
from typing import List
from typing import Mapping
from typing import Optional
from typing import Sequence
from typing import Set
from typing import Tuple
from typing import Type
from typing import TypeVar


py314b1 = sys.version_info >= (3, 14, 0, "beta", 1)
py314 = sys.version_info >= (3, 14)
py313 = sys.version_info >= (3, 13)
py312 = sys.version_info >= (3, 12)
py311 = sys.version_info >= (3, 11)
py310 = sys.version_info >= (3, 10)
py39 = sys.version_info >= (3, 9)
py38 = sys.version_info >= (3, 8)
pypy = platform.python_implementation() == "PyPy"
cpython = platform.python_implementation() == "CPython"

win32 = sys.platform.startswith("win")
osx = sys.platform.startswith("darwin")
arm = "aarch" in platform.machine().lower()
is64bit = sys.maxsize > 2**32

has_refcount_gc = bool(cpython)

dottedgetter = operator.attrgetter

_T_co = TypeVar("_T_co", covariant=True)


class FullArgSpec(typing.NamedTuple):
    args: List[str]
    varargs: Optional[str]
    varkw: Optional[str]
    defaults: Optional[Tuple[Any, ...]]
    kwonlyargs: List[str]
    kwonlydefaults: Optional[Dict[str, Any]]
    annotations: Dict[str, Any]


def inspect_getfullargspec(func: Callable[..., Any]) -> FullArgSpec:
    """Fully vendored version of getfullargspec from Python 3.3."""

    if inspect.ismethod(func):
        func = func.__func__
    if not inspect.isfunction(func):
        raise TypeError(f"{func!r} is not a Python function")

    co = func.__code__
    if not inspect.iscode(co):
        raise TypeError(f"{co!r} is not a code object")

    nargs = co.co_argcount
    names = co.co_varnames
    nkwargs = co.co_kwonlyargcount
    args = list(names[:nargs])
    kwonlyargs = list(names[nargs : nargs + nkwargs])

    nargs += nkwargs
    varargs = None
    if co.co_flags & inspect.CO_VARARGS:
        varargs = co.co_varnames[nargs]
        nargs = nargs + 1
    varkw = None
    if co.co_flags & inspect.CO_VARKEYWORDS:
        varkw = co.co_varnames[nargs]

    return FullArgSpec(
        args,
        varargs,
        varkw,
        func.__defaults__,
        kwonlyargs,
        func.__kwdefaults__,
        func.__annotations__,
    )


if py39:
    # python stubs don't have a public type for this. not worth
    # making a protocol
    def md5_not_for_security() -> Any:
        return hashlib.md5(usedforsecurity=False)

else:

    def md5_not_for_security() -> Any:
        return hashlib.md5()


if typing.TYPE_CHECKING or py38:
    from importlib import metadata as importlib_metadata
else:
    import importlib_metadata  # noqa


if typing.TYPE_CHECKING or py39:
    # pep 584 dict union
    dict_union = operator.or_  # noqa
else:

    def dict_union(a: dict, b: dict) -> dict:
        a = a.copy()
        a.update(b)
        return a


if py310:
    anext_ = anext
else:
    _NOT_PROVIDED = object()
    from collections.abc import AsyncIterator

    async def anext_(async_iterator, default=_NOT_PROVIDED):
        """vendored from https://github.com/python/cpython/pull/8895"""

        if not isinstance(async_iterator, AsyncIterator):
            raise TypeError(
                f"anext expected an AsyncIterator, got {type(async_iterator)}"
            )
        anxt = type(async_iterator).__anext__
        try:
            return await anxt(async_iterator)
        except StopAsyncIteration:
            if default is _NOT_PROVIDED:
                raise
            return default


def importlib_metadata_get(group):
    ep = importlib_metadata.entry_points()
    if typing.TYPE_CHECKING or hasattr(ep, "select"):
        return ep.select(group=group)
    else:
        return ep.get(group, ())


def b(s):
    return s.encode("latin-1")


def b64decode(x: str) -> bytes:
    return base64.b64decode(x.encode("ascii"))


def b64encode(x: bytes) -> str:
    return base64.b64encode(x).decode("ascii")


def decode_backslashreplace(text: bytes, encoding: str) -> str:
    return text.decode(encoding, errors="backslashreplace")


def cmp(a, b):
    return (a > b) - (a < b)


def _formatannotation(annotation, base_module=None):
    """vendored from python 3.7"""

    if isinstance(annotation, str):
        return annotation

    if getattr(annotation, "__module__", None) == "typing":
        return repr(annotation).replace("typing.", "").replace("~", "")
    if isinstance(annotation, type):
        if annotation.__module__ in ("builtins", base_module):
            return repr(annotation.__qualname__)
        return annotation.__module__ + "." + annotation.__qualname__
    elif isinstance(annotation, typing.TypeVar):
        return repr(annotation).replace("~", "")
    return repr(annotation).replace("~", "")


def inspect_formatargspec(
    args: List[str],
    varargs: Optional[str] = None,
    varkw: Optional[str] = None,
    defaults: Optional[Sequence[Any]] = None,
    kwonlyargs: Optional[Sequence[str]] = (),
    kwonlydefaults: Optional[Mapping[str, Any]] = {},
    annotations: Mapping[str, Any] = {},
    formatarg: Callable[[str], str] = str,
    formatvarargs: Callable[[str], str] = lambda name: "*" + name,
    formatvarkw: Callable[[str], str] = lambda name: "**" + name,
    formatvalue: Callable[[Any], str] = lambda value: "=" + repr(value),
    formatreturns: Callable[[Any], str] = lambda text: " -> " + str(text),
    formatannotation: Callable[[Any], str] = _formatannotation,
) -> str:
    """Copy formatargspec from python 3.7 standard library.

    Python 3 has deprecated formatargspec and requested that Signature
    be used instead, however this requires a full reimplementation
    of formatargspec() in terms of creating Parameter objects and such.
    Instead of introducing all the object-creation overhead and having
    to reinvent from scratch, just copy their compatibility routine.

    Ultimately we would need to rewrite our "decorator" routine completely
    which is not really worth it right now, until all Python 2.x support
    is dropped.

    """

    kwonlydefaults = kwonlydefaults or {}
    annotations = annotations or {}

    def formatargandannotation(arg):
        result = formatarg(arg)
        if arg in annotations:
            result += ": " + formatannotation(annotations[arg])
        return result

    specs = []
    if defaults:
        firstdefault = len(args) - len(defaults)
    else:
        firstdefault = -1

    for i, arg in enumerate(args):
        spec = formatargandannotation(arg)
        if defaults and i >= firstdefault:
            spec = spec + formatvalue(defaults[i - firstdefault])
        specs.append(spec)

    if varargs is not None:
        specs.append(formatvarargs(formatargandannotation(varargs)))
    else:
        if kwonlyargs:
            specs.append("*")

    if kwonlyargs:
        for kwonlyarg in kwonlyargs:
            spec = formatargandannotation(kwonlyarg)
            if kwonlydefaults and kwonlyarg in kwonlydefaults:
                spec += formatvalue(kwonlydefaults[kwonlyarg])
            specs.append(spec)

    if varkw is not None:
        specs.append(formatvarkw(formatargandannotation(varkw)))

    result = "(" + ", ".join(specs) + ")"
    if "return" in annotations:
        result += formatreturns(formatannotation(annotations["return"]))
    return result


def dataclass_fields(cls: Type[Any]) -> Iterable[dataclasses.Field[Any]]:
    """Return a sequence of all dataclasses.Field objects associated
    with a class as an already processed dataclass.

    The class must **already be a dataclass** for Field objects to be returned.

    """

    if dataclasses.is_dataclass(cls):
        return dataclasses.fields(cls)
    else:
        return []


def local_dataclass_fields(cls: Type[Any]) -> Iterable[dataclasses.Field[Any]]:
    """Return a sequence of all dataclasses.Field objects associated with
    an already processed dataclass, excluding those that originate from a
    superclass.

    The class must **already be a dataclass** for Field objects to be returned.

    """

    if dataclasses.is_dataclass(cls):
        super_fields: Set[dataclasses.Field[Any]] = set()
        for sup in cls.__bases__:
            super_fields.update(dataclass_fields(sup))
        return [f for f in dataclasses.fields(cls) if f not in super_fields]
    else:
        return []

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\expressions\permutation.py ===
from sympy.core import S
from sympy.core.sympify import _sympify
from sympy.functions import KroneckerDelta

from .matexpr import MatrixExpr
from .special import ZeroMatrix, Identity, OneMatrix


class PermutationMatrix(MatrixExpr):
    """A Permutation Matrix

    Parameters
    ==========

    perm : Permutation
        The permutation the matrix uses.

        The size of the permutation determines the matrix size.

        See the documentation of
        :class:`sympy.combinatorics.permutations.Permutation` for
        the further information of how to create a permutation object.

    Examples
    ========

    >>> from sympy import Matrix, PermutationMatrix
    >>> from sympy.combinatorics import Permutation

    Creating a permutation matrix:

    >>> p = Permutation(1, 2, 0)
    >>> P = PermutationMatrix(p)
    >>> P = P.as_explicit()
    >>> P
    Matrix([
    [0, 1, 0],
    [0, 0, 1],
    [1, 0, 0]])

    Permuting a matrix row and column:

    >>> M = Matrix([0, 1, 2])
    >>> Matrix(P*M)
    Matrix([
    [1],
    [2],
    [0]])

    >>> Matrix(M.T*P)
    Matrix([[2, 0, 1]])

    See Also
    ========

    sympy.combinatorics.permutations.Permutation
    """

    def __new__(cls, perm):
        from sympy.combinatorics.permutations import Permutation

        perm = _sympify(perm)
        if not isinstance(perm, Permutation):
            raise ValueError(
                "{} must be a SymPy Permutation instance.".format(perm))

        return super().__new__(cls, perm)

    @property
    def shape(self):
        size = self.args[0].size
        return (size, size)

    @property
    def is_Identity(self):
        return self.args[0].is_Identity

    def doit(self, **hints):
        if self.is_Identity:
            return Identity(self.rows)
        return self

    def _entry(self, i, j, **kwargs):
        perm = self.args[0]
        return KroneckerDelta(perm.apply(i), j)

    def _eval_power(self, exp):
        return PermutationMatrix(self.args[0] ** exp).doit()

    def _eval_inverse(self):
        return PermutationMatrix(self.args[0] ** -1)

    _eval_transpose = _eval_adjoint = _eval_inverse

    def _eval_determinant(self):
        sign = self.args[0].signature()
        if sign == 1:
            return S.One
        elif sign == -1:
            return S.NegativeOne
        raise NotImplementedError

    def _eval_rewrite_as_BlockDiagMatrix(self, *args, **kwargs):
        from sympy.combinatorics.permutations import Permutation
        from .blockmatrix import BlockDiagMatrix

        perm = self.args[0]
        full_cyclic_form = perm.full_cyclic_form

        cycles_picks = []

        # Stage 1. Decompose the cycles into the blockable form.
        a, b, c = 0, 0, 0
        flag = False
        for cycle in full_cyclic_form:
            l = len(cycle)
            m = max(cycle)

            if not flag:
                if m + 1 > a + l:
                    flag = True
                    temp = [cycle]
                    b = m
                    c = l
                else:
                    cycles_picks.append([cycle])
                    a += l

            else:
                if m > b:
                    if m + 1 == a + c + l:
                        temp.append(cycle)
                        cycles_picks.append(temp)
                        flag = False
                        a = m+1
                    else:
                        b = m
                        temp.append(cycle)
                        c += l
                else:
                    if b + 1 == a + c + l:
                        temp.append(cycle)
                        cycles_picks.append(temp)
                        flag = False
                        a = b+1
                    else:
                        temp.append(cycle)
                        c += l

        # Stage 2. Normalize each decomposed cycles and build matrix.
        p = 0
        args = []
        for pick in cycles_picks:
            new_cycles = []
            l = 0
            for cycle in pick:
                new_cycle = [i - p for i in cycle]
                new_cycles.append(new_cycle)
                l += len(cycle)
            p += l
            perm = Permutation(new_cycles)
            mat = PermutationMatrix(perm)
            args.append(mat)

        return BlockDiagMatrix(*args)


class MatrixPermute(MatrixExpr):
    r"""Symbolic representation for permuting matrix rows or columns.

    Parameters
    ==========

    perm : Permutation, PermutationMatrix
        The permutation to use for permuting the matrix.
        The permutation can be resized to the suitable one,

    axis : 0 or 1
        The axis to permute alongside.
        If `0`, it will permute the matrix rows.
        If `1`, it will permute the matrix columns.

    Notes
    =====

    This follows the same notation used in
    :meth:`sympy.matrices.matrixbase.MatrixBase.permute`.

    Examples
    ========

    >>> from sympy import Matrix, MatrixPermute
    >>> from sympy.combinatorics import Permutation

    Permuting the matrix rows:

    >>> p = Permutation(1, 2, 0)
    >>> A = Matrix([[1, 2, 3], [4, 5, 6], [7, 8, 9]])
    >>> B = MatrixPermute(A, p, axis=0)
    >>> B.as_explicit()
    Matrix([
    [4, 5, 6],
    [7, 8, 9],
    [1, 2, 3]])

    Permuting the matrix columns:

    >>> B = MatrixPermute(A, p, axis=1)
    >>> B.as_explicit()
    Matrix([
    [2, 3, 1],
    [5, 6, 4],
    [8, 9, 7]])

    See Also
    ========

    sympy.matrices.matrixbase.MatrixBase.permute
    """
    def __new__(cls, mat, perm, axis=S.Zero):
        from sympy.combinatorics.permutations import Permutation

        mat = _sympify(mat)
        if not mat.is_Matrix:
            raise ValueError(
                "{} must be a SymPy matrix instance.".format(perm))

        perm = _sympify(perm)
        if isinstance(perm, PermutationMatrix):
            perm = perm.args[0]

        if not isinstance(perm, Permutation):
            raise ValueError(
                "{} must be a SymPy Permutation or a PermutationMatrix " \
                "instance".format(perm))

        axis = _sympify(axis)
        if axis not in (0, 1):
            raise ValueError("The axis must be 0 or 1.")

        mat_size = mat.shape[axis]
        if mat_size != perm.size:
            try:
                perm = perm.resize(mat_size)
            except ValueError:
                raise ValueError(
                    "Size does not match between the permutation {} "
                    "and the matrix {} threaded over the axis {} "
                    "and cannot be converted."
                    .format(perm, mat, axis))

        return super().__new__(cls, mat, perm, axis)

    def doit(self, deep=True, **hints):
        mat, perm, axis = self.args

        if deep:
            mat = mat.doit(deep=deep, **hints)
            perm = perm.doit(deep=deep, **hints)

        if perm.is_Identity:
            return mat

        if mat.is_Identity:
            if axis is S.Zero:
                return PermutationMatrix(perm)
            elif axis is S.One:
                return PermutationMatrix(perm**-1)

        if isinstance(mat, (ZeroMatrix, OneMatrix)):
            return mat

        if isinstance(mat, MatrixPermute) and mat.args[2] == axis:
            return MatrixPermute(mat.args[0], perm * mat.args[1], axis)

        return self

    @property
    def shape(self):
        return self.args[0].shape

    def _entry(self, i, j, **kwargs):
        mat, perm, axis = self.args

        if axis == 0:
            return mat[perm.apply(i), j]
        elif axis == 1:
            return mat[i, perm.apply(j)]

    def _eval_rewrite_as_MatMul(self, *args, **kwargs):
        from .matmul import MatMul

        mat, perm, axis = self.args

        deep = kwargs.get("deep", True)

        if deep:
            mat = mat.rewrite(MatMul)

        if axis == 0:
            return MatMul(PermutationMatrix(perm), mat)
        elif axis == 1:
            return MatMul(mat, PermutationMatrix(perm**-1))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\t5\t5_helper.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation. All rights reserved.
# Licensed under the MIT License.
# -------------------------------------------------------------------------

import logging
import os
from pathlib import Path

import torch
from float16 import float_to_float16_max_diff
from onnx_model import OnnxModel
from optimizer import optimize_model
from t5_decoder import T5Decoder, T5DecoderHelper
from t5_encoder_decoder_init import T5EncoderDecoderInit, T5EncoderDecoderInitHelper
from transformers import MT5ForConditionalGeneration, T5ForConditionalGeneration

from onnxruntime import InferenceSession

logger = logging.getLogger(__name__)

PRETRAINED_T5_MODELS = ["t5-small", "t5-base", "t5-large", "t5-3b", "t5-11b"]
PRETRAINED_MT5_MODELS = [
    "google/mt5-small",
    "google/mt5-base",
    "google/mt5-large",
    "google/mt5-xl",
    "google/mt5-xxl",
]


class T5Helper:
    @staticmethod
    def get_onnx_path(
        output_dir: str,
        model_name_or_path: str,
        suffix: str = "",
        new_folder: bool = False,
    ) -> str:
        """Build onnx path

        Args:
            output_dir (str): output directory
            model_name_or_path (str): pretrained model name, or path to the model checkpoint
            suffix (str, optional): suffix like "_encoder" or "_decoder_fp16" will be appended to file name. Defaults to None.
            new_folder (bool, optional): create a new directory for the model. Defaults to False.

        Returns:
            str: path of onnx model
        """
        model_name = model_name_or_path
        if os.path.isdir(model_name_or_path):
            model_name = Path(model_name_or_path).parts[-1]
        else:
            model_name.split("/")[-1]

        model_name += suffix

        directory = os.path.join(output_dir, model_name) if new_folder else output_dir
        return os.path.join(directory, model_name + ".onnx")

    @staticmethod
    def load_model(
        model_name_or_path: str,
        cache_dir: str,
        device: torch.device,
        model_type: str = "t5",
        state_dict_path: str = "",
        encoder_decoder_init: bool = False,
    ) -> dict[str, T5EncoderDecoderInit | T5Decoder]:
        """Load model given a pretrained name or path, then build models for ONNX conversion.

        Args:
            model_name_or_path (str): pretrained model name or path
            cache_dir (str): cache directory
            device (torch.device): device to run the model
            model_type (str, optional): model type "t5" or "mt5"
            state_dict_path(str, optional): state dictionary path
            encoder_decoder_init (bool, optional): combine encoder and decoder kv cache initialization into one model.
        Returns:
            Dict[str, torch.nn.Module]: mapping from name to modules for ONNX conversion.
        """
        if model_type == "t5":
            model = T5ForConditionalGeneration.from_pretrained(model_name_or_path, cache_dir=cache_dir)
        elif model_type == "mt5":
            model = MT5ForConditionalGeneration.from_pretrained(model_name_or_path, cache_dir=cache_dir)
        else:
            raise ValueError("only support mode_type=t5 or mt5")

        if state_dict_path:
            model.load_state_dict(torch.load(state_dict_path))

        decoder = T5Decoder(model.decoder, model.lm_head, model.config)
        decoder.eval().to(device)

        encoder = T5EncoderDecoderInit(
            model.encoder,
            model.decoder,
            model.lm_head,
            model.config,
            decoder_start_token_id=None,
            output_cross_only=not encoder_decoder_init,
        )

        encoder_name = "encoder_decoder_init" if encoder_decoder_init else "encoder"
        return {encoder_name: encoder, "decoder": decoder}

    @staticmethod
    def export_onnx(
        model: T5Decoder | T5EncoderDecoderInit,
        device: torch.device,
        onnx_model_path: str,
        verbose: bool = True,
        use_external_data_format: bool = False,
        use_decoder_input_ids: bool = True,
        use_int32_inputs: bool = False,
    ):
        if isinstance(model, T5EncoderDecoderInit):
            T5EncoderDecoderInitHelper.export_onnx(
                model,
                device,
                onnx_model_path,
                use_decoder_input_ids,
                verbose,
                use_external_data_format,
                use_int32_inputs,
            )
        else:
            T5DecoderHelper.export_onnx(
                model,
                device,
                onnx_model_path,
                verbose,
                use_external_data_format,
                use_int32_inputs,
            )

    @staticmethod
    def auto_mixed_precision(
        onnx_model: OnnxModel,
        op_block_list: list[str] | None = None,
        force_fp16_logits: bool = False,
        use_symbolic_shape_infer: bool = True,
    ):
        """Convert model to mixed precision.
           It detects whether original model has fp16 precision weights, and set parameters for float16 conversion automatically.
        Args:
            onnx_model (OnnxModel): optimized ONNX model
            op_block_list (List[str], optional): operators need to run in fp32.
            force_fp16_logits (bool, optional): force logits and last MatMul node to be in float16. Defaults to False.
            use_symbolic_shape_infer (bool, optional): use symbolic shape inference to convert float to float16. Defaults to True.
        Returns:
            parameters(dict): a dictionary of parameters used in float16 conversion
        """
        if op_block_list is None:
            op_block_list = [
                "SimplifiedLayerNormalization",
                "SkipSimplifiedLayerNormalization",
                "Relu",
                "Add",
            ]

        op_full_set = {node.op_type for node in onnx_model.nodes()}
        fp32_op_set = set(op_block_list)
        fp16_op_set = op_full_set.difference(fp32_op_set)
        logger.info(f"fp32 op: {fp32_op_set} fp16 op: {fp16_op_set}")

        # logits is the first output
        logits_output_name = onnx_model.graph().output[0].name

        # We use the weight in last MatMul node to detect whether the model is stored with float16 weights from training.
        is_weight_fp16_precision = False
        output_name_to_node = onnx_model.output_name_to_node()
        assert logits_output_name in output_name_to_node
        node = output_name_to_node[logits_output_name]
        last_matmul_node = None
        if node.op_type == "MatMul":
            last_matmul_node = node
            logger.info(f"Found last MatMul node for logits: {node.name}")
            initializer = None
            for input in node.input:
                initializer = onnx_model.get_initializer(input)
                if initializer is not None:
                    break

            # when the max difference of value after converting float to float16 is lower than a threshold (1e-6),
            # we can deduce that the weights are stored in float16 precision.
            max_diff = float_to_float16_max_diff(initializer)
            logger.debug(f"max diff of converting weights in last MatMul node {node.name}: {max_diff}")
            is_weight_fp16_precision = max_diff < 1e-6
        else:
            logger.warning(f"Failed to find MatMul node for logits. Found {node.op_type} of node {node.name}")

        keep_io_types = []
        node_block_list = []
        if (not is_weight_fp16_precision) and (last_matmul_node is not None) and not force_fp16_logits:
            # When original weight is float32 precision, keep logits and last MatMul in float32 could get better precision.
            keep_io_types = [logits_output_name]
            node_block_list = [last_matmul_node.name]

        if "Add" not in op_block_list:
            input_name_to_nodes = onnx_model.input_name_to_nodes()
            fp32_add = 0
            changed = True
            add_nodes = onnx_model.get_nodes_by_op_type("Add")
            while changed:
                changed = False
                for node in add_nodes:
                    if node.name not in node_block_list:
                        parents = onnx_model.get_parents(node, output_name_to_node)
                        children = onnx_model.get_children(node, input_name_to_nodes)
                        blocked_children = [
                            child for child in children if child.op_type in op_block_list or child in node_block_list
                        ]
                        blocked_parents = [
                            parent for parent in parents if parent.op_type in op_block_list or parent in node_block_list
                        ]
                        # If any child or parent is in fp32, we place the Add node to fp32.
                        if (len(blocked_children) + len(blocked_parents)) > 0:
                            node_block_list.append(node.name)
                            fp32_add += 1
                            changed = True
            fp16_add = len(add_nodes) - fp32_add
            logger.info(f"node counter of Add operator: fp32={fp32_add} fp16={fp16_add}")

        logger.info(f"node_block_list: {node_block_list}")

        parameters = {
            "keep_io_types": keep_io_types,
            "op_block_list": op_block_list,
            "node_block_list": node_block_list,
            "force_fp16_initializers": is_weight_fp16_precision,
        }

        logger.info(f"auto_mixed_precision parameters: {parameters}")
        if use_symbolic_shape_infer:
            onnx_model.convert_float_to_float16(use_symbolic_shape_infer=True, **parameters)
        else:
            # Workaround when symbolic shape inference fails.
            # Need enable shape_infer_before_optimization in convert_to_onnx.py as well.
            from float16 import convert_float_to_float16

            convert_float_to_float16(
                onnx_model.model,
                disable_shape_infer=True,
                **parameters,
            )

        return parameters

    @staticmethod
    def optimize_onnx(
        onnx_model_path: str,
        optimized_model_path: str,
        is_float16: bool,
        num_attention_heads: int,
        hidden_size: int,
        use_external_data_format: bool = False,
        auto_mixed_precision: bool = True,
        use_gpu: bool = False,
        force_fp16_io: bool = False,
    ):
        """Optimize ONNX model with an option to convert it to use mixed precision."""

        from fusion_options import FusionOptions

        optimization_options = None
        if is_float16:
            optimization_options = FusionOptions("t5")
            # SkipLayerNormalization is faster but might bring accuracy drop since it uses fp16 accumulation.
            optimization_options.enable_skip_layer_norm = not auto_mixed_precision

        m = optimize_model(
            onnx_model_path,
            model_type="t5",
            num_heads=num_attention_heads,
            hidden_size=hidden_size,
            opt_level=0,
            optimization_options=optimization_options,
            use_gpu=use_gpu,
        )

        if is_float16:
            if auto_mixed_precision:
                T5Helper.auto_mixed_precision(m, force_fp16_logits=force_fp16_io)
            else:
                m.convert_model_float32_to_float16(cast_input_output=force_fp16_io)

        m.save_model_to_file(optimized_model_path, use_external_data_format, all_tensors_to_one_file=True)

    @staticmethod
    def verify_onnx(
        model: T5Decoder | T5EncoderDecoderInit,
        ort_session: InferenceSession,
        device: torch.device,
        use_int32_inputs: bool,
    ):
        """Compare the result from PyTorch and OnnxRuntime to verify the ONNX model is good."""
        if isinstance(model, T5EncoderDecoderInit):
            return T5EncoderDecoderInitHelper.verify_onnx(model, ort_session, device, use_int32_inputs)

        return T5DecoderHelper.verify_onnx(model, ort_session, device, use_int32_inputs)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\orm\identity.py ===
# orm/identity.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

from __future__ import annotations

from typing import Any
from typing import cast
from typing import Dict
from typing import Iterable
from typing import Iterator
from typing import List
from typing import NoReturn
from typing import Optional
from typing import Set
from typing import Tuple
from typing import TYPE_CHECKING
from typing import TypeVar
import weakref

from . import util as orm_util
from .. import exc as sa_exc

if TYPE_CHECKING:
    from ._typing import _IdentityKeyType
    from .state import InstanceState


_T = TypeVar("_T", bound=Any)

_O = TypeVar("_O", bound=object)


class IdentityMap:
    _wr: weakref.ref[IdentityMap]

    _dict: Dict[_IdentityKeyType[Any], Any]
    _modified: Set[InstanceState[Any]]

    def __init__(self) -> None:
        self._dict = {}
        self._modified = set()
        self._wr = weakref.ref(self)

    def _kill(self) -> None:
        self._add_unpresent = _killed  # type: ignore

    def all_states(self) -> List[InstanceState[Any]]:
        raise NotImplementedError()

    def contains_state(self, state: InstanceState[Any]) -> bool:
        raise NotImplementedError()

    def __contains__(self, key: _IdentityKeyType[Any]) -> bool:
        raise NotImplementedError()

    def safe_discard(self, state: InstanceState[Any]) -> None:
        raise NotImplementedError()

    def __getitem__(self, key: _IdentityKeyType[_O]) -> _O:
        raise NotImplementedError()

    def get(
        self, key: _IdentityKeyType[_O], default: Optional[_O] = None
    ) -> Optional[_O]:
        raise NotImplementedError()

    def fast_get_state(
        self, key: _IdentityKeyType[_O]
    ) -> Optional[InstanceState[_O]]:
        raise NotImplementedError()

    def keys(self) -> Iterable[_IdentityKeyType[Any]]:
        return self._dict.keys()

    def values(self) -> Iterable[object]:
        raise NotImplementedError()

    def replace(self, state: InstanceState[_O]) -> Optional[InstanceState[_O]]:
        raise NotImplementedError()

    def add(self, state: InstanceState[Any]) -> bool:
        raise NotImplementedError()

    def _fast_discard(self, state: InstanceState[Any]) -> None:
        raise NotImplementedError()

    def _add_unpresent(
        self, state: InstanceState[Any], key: _IdentityKeyType[Any]
    ) -> None:
        """optional inlined form of add() which can assume item isn't present
        in the map"""
        self.add(state)

    def _manage_incoming_state(self, state: InstanceState[Any]) -> None:
        state._instance_dict = self._wr

        if state.modified:
            self._modified.add(state)

    def _manage_removed_state(self, state: InstanceState[Any]) -> None:
        del state._instance_dict
        if state.modified:
            self._modified.discard(state)

    def _dirty_states(self) -> Set[InstanceState[Any]]:
        return self._modified

    def check_modified(self) -> bool:
        """return True if any InstanceStates present have been marked
        as 'modified'.

        """
        return bool(self._modified)

    def has_key(self, key: _IdentityKeyType[Any]) -> bool:
        return key in self

    def __len__(self) -> int:
        return len(self._dict)


class WeakInstanceDict(IdentityMap):
    _dict: Dict[_IdentityKeyType[Any], InstanceState[Any]]

    def __getitem__(self, key: _IdentityKeyType[_O]) -> _O:
        state = cast("InstanceState[_O]", self._dict[key])
        o = state.obj()
        if o is None:
            raise KeyError(key)
        return o

    def __contains__(self, key: _IdentityKeyType[Any]) -> bool:
        try:
            if key in self._dict:
                state = self._dict[key]
                o = state.obj()
            else:
                return False
        except KeyError:
            return False
        else:
            return o is not None

    def contains_state(self, state: InstanceState[Any]) -> bool:
        if state.key in self._dict:
            if TYPE_CHECKING:
                assert state.key is not None
            try:
                return self._dict[state.key] is state
            except KeyError:
                return False
        else:
            return False

    def replace(
        self, state: InstanceState[Any]
    ) -> Optional[InstanceState[Any]]:
        assert state.key is not None
        if state.key in self._dict:
            try:
                existing = existing_non_none = self._dict[state.key]
            except KeyError:
                # catch gc removed the key after we just checked for it
                existing = None
            else:
                if existing_non_none is not state:
                    self._manage_removed_state(existing_non_none)
                else:
                    return None
        else:
            existing = None

        self._dict[state.key] = state
        self._manage_incoming_state(state)
        return existing

    def add(self, state: InstanceState[Any]) -> bool:
        key = state.key
        assert key is not None
        # inline of self.__contains__
        if key in self._dict:
            try:
                existing_state = self._dict[key]
            except KeyError:
                # catch gc removed the key after we just checked for it
                pass
            else:
                if existing_state is not state:
                    o = existing_state.obj()
                    if o is not None:
                        raise sa_exc.InvalidRequestError(
                            "Can't attach instance "
                            "%s; another instance with key %s is already "
                            "present in this session."
                            % (orm_util.state_str(state), state.key)
                        )
                else:
                    return False
        self._dict[key] = state
        self._manage_incoming_state(state)
        return True

    def _add_unpresent(
        self, state: InstanceState[Any], key: _IdentityKeyType[Any]
    ) -> None:
        # inlined form of add() called by loading.py
        self._dict[key] = state
        state._instance_dict = self._wr

    def fast_get_state(
        self, key: _IdentityKeyType[_O]
    ) -> Optional[InstanceState[_O]]:
        return self._dict.get(key)

    def get(
        self, key: _IdentityKeyType[_O], default: Optional[_O] = None
    ) -> Optional[_O]:
        if key not in self._dict:
            return default
        try:
            state = cast("InstanceState[_O]", self._dict[key])
        except KeyError:
            # catch gc removed the key after we just checked for it
            return default
        else:
            o = state.obj()
            if o is None:
                return default
            return o

    def items(self) -> List[Tuple[_IdentityKeyType[Any], InstanceState[Any]]]:
        values = self.all_states()
        result = []
        for state in values:
            value = state.obj()
            key = state.key
            assert key is not None
            if value is not None:
                result.append((key, value))
        return result

    def values(self) -> List[object]:
        values = self.all_states()
        result = []
        for state in values:
            value = state.obj()
            if value is not None:
                result.append(value)

        return result

    def __iter__(self) -> Iterator[_IdentityKeyType[Any]]:
        return iter(self.keys())

    def all_states(self) -> List[InstanceState[Any]]:
        return list(self._dict.values())

    def _fast_discard(self, state: InstanceState[Any]) -> None:
        # used by InstanceState for state being
        # GC'ed, inlines _managed_removed_state
        key = state.key
        assert key is not None
        try:
            st = self._dict[key]
        except KeyError:
            # catch gc removed the key after we just checked for it
            pass
        else:
            if st is state:
                self._dict.pop(key, None)

    def discard(self, state: InstanceState[Any]) -> None:
        self.safe_discard(state)

    def safe_discard(self, state: InstanceState[Any]) -> None:
        key = state.key
        if key in self._dict:
            assert key is not None
            try:
                st = self._dict[key]
            except KeyError:
                # catch gc removed the key after we just checked for it
                pass
            else:
                if st is state:
                    self._dict.pop(key, None)
                    self._manage_removed_state(state)


def _killed(state: InstanceState[Any], key: _IdentityKeyType[Any]) -> NoReturn:
    # external function to avoid creating cycles when assigned to
    # the IdentityMap
    raise sa_exc.InvalidRequestError(
        "Object %s cannot be converted to 'persistent' state, as this "
        "identity map is no longer valid.  Has the owning Session "
        "been closed?" % orm_util.state_str(state),
        code="lkrp",
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\assumptions\relation\equality.py ===
"""
Module for mathematical equality [1] and inequalities [2].

The purpose of this module is to provide the instances which represent the
binary predicates in order to combine the relationals into logical inference
system. Objects such as ``Q.eq``, ``Q.lt`` should remain internal to
assumptions module, and user must use the classes such as :obj:`~.Eq()`,
:obj:`~.Lt()` instead to construct the relational expressions.

References
==========

.. [1] https://en.wikipedia.org/wiki/Equality_(mathematics)
.. [2] https://en.wikipedia.org/wiki/Inequality_(mathematics)
"""
from sympy.assumptions import Q
from sympy.core.relational import is_eq, is_neq, is_gt, is_ge, is_lt, is_le

from .binrel import BinaryRelation

__all__ = ['EqualityPredicate', 'UnequalityPredicate', 'StrictGreaterThanPredicate',
    'GreaterThanPredicate', 'StrictLessThanPredicate', 'LessThanPredicate']


class EqualityPredicate(BinaryRelation):
    """
    Binary predicate for $=$.

    The purpose of this class is to provide the instance which represent
    the equality predicate in order to allow the logical inference.
    This class must remain internal to assumptions module and user must
    use :obj:`~.Eq()` instead to construct the equality expression.

    Evaluating this predicate to ``True`` or ``False`` is done by
    :func:`~.core.relational.is_eq`

    Examples
    ========

    >>> from sympy import ask, Q
    >>> Q.eq(0, 0)
    Q.eq(0, 0)
    >>> ask(_)
    True

    See Also
    ========

    sympy.core.relational.Eq

    """
    is_reflexive = True
    is_symmetric = True

    name = 'eq'
    handler = None  # Do not allow dispatching by this predicate

    @property
    def negated(self):
        return Q.ne

    def eval(self, args, assumptions=True):
        if assumptions == True:
            # default assumptions for is_eq is None
            assumptions = None
        return is_eq(*args, assumptions)


class UnequalityPredicate(BinaryRelation):
    r"""
    Binary predicate for $\neq$.

    The purpose of this class is to provide the instance which represent
    the inequation predicate in order to allow the logical inference.
    This class must remain internal to assumptions module and user must
    use :obj:`~.Ne()` instead to construct the inequation expression.

    Evaluating this predicate to ``True`` or ``False`` is done by
    :func:`~.core.relational.is_neq`

    Examples
    ========

    >>> from sympy import ask, Q
    >>> Q.ne(0, 0)
    Q.ne(0, 0)
    >>> ask(_)
    False

    See Also
    ========

    sympy.core.relational.Ne

    """
    is_reflexive = False
    is_symmetric = True

    name = 'ne'
    handler = None

    @property
    def negated(self):
        return Q.eq

    def eval(self, args, assumptions=True):
        if assumptions == True:
            # default assumptions for is_neq is None
            assumptions = None
        return is_neq(*args, assumptions)


class StrictGreaterThanPredicate(BinaryRelation):
    """
    Binary predicate for $>$.

    The purpose of this class is to provide the instance which represent
    the ">" predicate in order to allow the logical inference.
    This class must remain internal to assumptions module and user must
    use :obj:`~.Gt()` instead to construct the equality expression.

    Evaluating this predicate to ``True`` or ``False`` is done by
    :func:`~.core.relational.is_gt`

    Examples
    ========

    >>> from sympy import ask, Q
    >>> Q.gt(0, 0)
    Q.gt(0, 0)
    >>> ask(_)
    False

    See Also
    ========

    sympy.core.relational.Gt

    """
    is_reflexive = False
    is_symmetric = False

    name = 'gt'
    handler = None

    @property
    def reversed(self):
        return Q.lt

    @property
    def negated(self):
        return Q.le

    def eval(self, args, assumptions=True):
        if assumptions == True:
            # default assumptions for is_gt is None
            assumptions = None
        return is_gt(*args, assumptions)


class GreaterThanPredicate(BinaryRelation):
    """
    Binary predicate for $>=$.

    The purpose of this class is to provide the instance which represent
    the ">=" predicate in order to allow the logical inference.
    This class must remain internal to assumptions module and user must
    use :obj:`~.Ge()` instead to construct the equality expression.

    Evaluating this predicate to ``True`` or ``False`` is done by
    :func:`~.core.relational.is_ge`

    Examples
    ========

    >>> from sympy import ask, Q
    >>> Q.ge(0, 0)
    Q.ge(0, 0)
    >>> ask(_)
    True

    See Also
    ========

    sympy.core.relational.Ge

    """
    is_reflexive = True
    is_symmetric = False

    name = 'ge'
    handler = None

    @property
    def reversed(self):
        return Q.le

    @property
    def negated(self):
        return Q.lt

    def eval(self, args, assumptions=True):
        if assumptions == True:
            # default assumptions for is_ge is None
            assumptions = None
        return is_ge(*args, assumptions)


class StrictLessThanPredicate(BinaryRelation):
    """
    Binary predicate for $<$.

    The purpose of this class is to provide the instance which represent
    the "<" predicate in order to allow the logical inference.
    This class must remain internal to assumptions module and user must
    use :obj:`~.Lt()` instead to construct the equality expression.

    Evaluating this predicate to ``True`` or ``False`` is done by
    :func:`~.core.relational.is_lt`

    Examples
    ========

    >>> from sympy import ask, Q
    >>> Q.lt(0, 0)
    Q.lt(0, 0)
    >>> ask(_)
    False

    See Also
    ========

    sympy.core.relational.Lt

    """
    is_reflexive = False
    is_symmetric = False

    name = 'lt'
    handler = None

    @property
    def reversed(self):
        return Q.gt

    @property
    def negated(self):
        return Q.ge

    def eval(self, args, assumptions=True):
        if assumptions == True:
            # default assumptions for is_lt is None
            assumptions = None
        return is_lt(*args, assumptions)


class LessThanPredicate(BinaryRelation):
    """
    Binary predicate for $<=$.

    The purpose of this class is to provide the instance which represent
    the "<=" predicate in order to allow the logical inference.
    This class must remain internal to assumptions module and user must
    use :obj:`~.Le()` instead to construct the equality expression.

    Evaluating this predicate to ``True`` or ``False`` is done by
    :func:`~.core.relational.is_le`

    Examples
    ========

    >>> from sympy import ask, Q
    >>> Q.le(0, 0)
    Q.le(0, 0)
    >>> ask(_)
    True

    See Also
    ========

    sympy.core.relational.Le

    """
    is_reflexive = True
    is_symmetric = False

    name = 'le'
    handler = None

    @property
    def reversed(self):
        return Q.ge

    @property
    def negated(self):
        return Q.gt

    def eval(self, args, assumptions=True):
        if assumptions == True:
            # default assumptions for is_le is None
            assumptions = None
        return is_le(*args, assumptions)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\click\formatting.py ===
from __future__ import annotations

import collections.abc as cabc
from contextlib import contextmanager
from gettext import gettext as _

from ._compat import term_len
from .parser import _split_opt

# Can force a width.  This is used by the test system
FORCED_WIDTH: int | None = None


def measure_table(rows: cabc.Iterable[tuple[str, str]]) -> tuple[int, ...]:
    widths: dict[int, int] = {}

    for row in rows:
        for idx, col in enumerate(row):
            widths[idx] = max(widths.get(idx, 0), term_len(col))

    return tuple(y for x, y in sorted(widths.items()))


def iter_rows(
    rows: cabc.Iterable[tuple[str, str]], col_count: int
) -> cabc.Iterator[tuple[str, ...]]:
    for row in rows:
        yield row + ("",) * (col_count - len(row))


def wrap_text(
    text: str,
    width: int = 78,
    initial_indent: str = "",
    subsequent_indent: str = "",
    preserve_paragraphs: bool = False,
) -> str:
    """A helper function that intelligently wraps text.  By default, it
    assumes that it operates on a single paragraph of text but if the
    `preserve_paragraphs` parameter is provided it will intelligently
    handle paragraphs (defined by two empty lines).

    If paragraphs are handled, a paragraph can be prefixed with an empty
    line containing the ``\\b`` character (``\\x08``) to indicate that
    no rewrapping should happen in that block.

    :param text: the text that should be rewrapped.
    :param width: the maximum width for the text.
    :param initial_indent: the initial indent that should be placed on the
                           first line as a string.
    :param subsequent_indent: the indent string that should be placed on
                              each consecutive line.
    :param preserve_paragraphs: if this flag is set then the wrapping will
                                intelligently handle paragraphs.
    """
    from ._textwrap import TextWrapper

    text = text.expandtabs()
    wrapper = TextWrapper(
        width,
        initial_indent=initial_indent,
        subsequent_indent=subsequent_indent,
        replace_whitespace=False,
    )
    if not preserve_paragraphs:
        return wrapper.fill(text)

    p: list[tuple[int, bool, str]] = []
    buf: list[str] = []
    indent = None

    def _flush_par() -> None:
        if not buf:
            return
        if buf[0].strip() == "\b":
            p.append((indent or 0, True, "\n".join(buf[1:])))
        else:
            p.append((indent or 0, False, " ".join(buf)))
        del buf[:]

    for line in text.splitlines():
        if not line:
            _flush_par()
            indent = None
        else:
            if indent is None:
                orig_len = term_len(line)
                line = line.lstrip()
                indent = orig_len - term_len(line)
            buf.append(line)
    _flush_par()

    rv = []
    for indent, raw, text in p:
        with wrapper.extra_indent(" " * indent):
            if raw:
                rv.append(wrapper.indent_only(text))
            else:
                rv.append(wrapper.fill(text))

    return "\n\n".join(rv)


class HelpFormatter:
    """This class helps with formatting text-based help pages.  It's
    usually just needed for very special internal cases, but it's also
    exposed so that developers can write their own fancy outputs.

    At present, it always writes into memory.

    :param indent_increment: the additional increment for each level.
    :param width: the width for the text.  This defaults to the terminal
                  width clamped to a maximum of 78.
    """

    def __init__(
        self,
        indent_increment: int = 2,
        width: int | None = None,
        max_width: int | None = None,
    ) -> None:
        import shutil

        self.indent_increment = indent_increment
        if max_width is None:
            max_width = 80
        if width is None:
            width = FORCED_WIDTH
            if width is None:
                width = max(min(shutil.get_terminal_size().columns, max_width) - 2, 50)
        self.width = width
        self.current_indent: int = 0
        self.buffer: list[str] = []

    def write(self, string: str) -> None:
        """Writes a unicode string into the internal buffer."""
        self.buffer.append(string)

    def indent(self) -> None:
        """Increases the indentation."""
        self.current_indent += self.indent_increment

    def dedent(self) -> None:
        """Decreases the indentation."""
        self.current_indent -= self.indent_increment

    def write_usage(self, prog: str, args: str = "", prefix: str | None = None) -> None:
        """Writes a usage line into the buffer.

        :param prog: the program name.
        :param args: whitespace separated list of arguments.
        :param prefix: The prefix for the first line. Defaults to
            ``"Usage: "``.
        """
        if prefix is None:
            prefix = f"{_('Usage:')} "

        usage_prefix = f"{prefix:>{self.current_indent}}{prog} "
        text_width = self.width - self.current_indent

        if text_width >= (term_len(usage_prefix) + 20):
            # The arguments will fit to the right of the prefix.
            indent = " " * term_len(usage_prefix)
            self.write(
                wrap_text(
                    args,
                    text_width,
                    initial_indent=usage_prefix,
                    subsequent_indent=indent,
                )
            )
        else:
            # The prefix is too long, put the arguments on the next line.
            self.write(usage_prefix)
            self.write("\n")
            indent = " " * (max(self.current_indent, term_len(prefix)) + 4)
            self.write(
                wrap_text(
                    args, text_width, initial_indent=indent, subsequent_indent=indent
                )
            )

        self.write("\n")

    def write_heading(self, heading: str) -> None:
        """Writes a heading into the buffer."""
        self.write(f"{'':>{self.current_indent}}{heading}:\n")

    def write_paragraph(self) -> None:
        """Writes a paragraph into the buffer."""
        if self.buffer:
            self.write("\n")

    def write_text(self, text: str) -> None:
        """Writes re-indented text into the buffer.  This rewraps and
        preserves paragraphs.
        """
        indent = " " * self.current_indent
        self.write(
            wrap_text(
                text,
                self.width,
                initial_indent=indent,
                subsequent_indent=indent,
                preserve_paragraphs=True,
            )
        )
        self.write("\n")

    def write_dl(
        self,
        rows: cabc.Sequence[tuple[str, str]],
        col_max: int = 30,
        col_spacing: int = 2,
    ) -> None:
        """Writes a definition list into the buffer.  This is how options
        and commands are usually formatted.

        :param rows: a list of two item tuples for the terms and values.
        :param col_max: the maximum width of the first column.
        :param col_spacing: the number of spaces between the first and
                            second column.
        """
        rows = list(rows)
        widths = measure_table(rows)
        if len(widths) != 2:
            raise TypeError("Expected two columns for definition list")

        first_col = min(widths[0], col_max) + col_spacing

        for first, second in iter_rows(rows, len(widths)):
            self.write(f"{'':>{self.current_indent}}{first}")
            if not second:
                self.write("\n")
                continue
            if term_len(first) <= first_col - col_spacing:
                self.write(" " * (first_col - term_len(first)))
            else:
                self.write("\n")
                self.write(" " * (first_col + self.current_indent))

            text_width = max(self.width - first_col - 2, 10)
            wrapped_text = wrap_text(second, text_width, preserve_paragraphs=True)
            lines = wrapped_text.splitlines()

            if lines:
                self.write(f"{lines[0]}\n")

                for line in lines[1:]:
                    self.write(f"{'':>{first_col + self.current_indent}}{line}\n")
            else:
                self.write("\n")

    @contextmanager
    def section(self, name: str) -> cabc.Iterator[None]:
        """Helpful context manager that writes a paragraph, a heading,
        and the indents.

        :param name: the section name that is written as heading.
        """
        self.write_paragraph()
        self.write_heading(name)
        self.indent()
        try:
            yield
        finally:
            self.dedent()

    @contextmanager
    def indentation(self) -> cabc.Iterator[None]:
        """A context manager that increases the indentation."""
        self.indent()
        try:
            yield
        finally:
            self.dedent()

    def getvalue(self) -> str:
        """Returns the buffer contents."""
        return "".join(self.buffer)


def join_options(options: cabc.Sequence[str]) -> tuple[str, bool]:
    """Given a list of option strings this joins them in the most appropriate
    way and returns them in the form ``(formatted_string,
    any_prefix_is_slash)`` where the second item in the tuple is a flag that
    indicates if any of the option prefixes was a slash.
    """
    rv = []
    any_prefix_is_slash = False

    for opt in options:
        prefix = _split_opt(opt)[0]

        if prefix == "/":
            any_prefix_is_slash = True

        rv.append((len(prefix), opt))

    rv.sort(key=lambda x: x[0])
    return ", ".join(x[1] for x in rv), any_prefix_is_slash

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\dtypes\astype.py ===
"""
Functions for implementing 'astype' methods according to pandas conventions,
particularly ones that differ from numpy.
"""
from __future__ import annotations

import inspect
from typing import (
    TYPE_CHECKING,
    overload,
)
import warnings

import numpy as np

from pandas._libs import lib
from pandas._libs.tslibs.timedeltas import array_to_timedelta64
from pandas.errors import IntCastingNaNError

from pandas.core.dtypes.common import (
    is_object_dtype,
    is_string_dtype,
    pandas_dtype,
)
from pandas.core.dtypes.dtypes import (
    ExtensionDtype,
    NumpyEADtype,
)

if TYPE_CHECKING:
    from pandas._typing import (
        ArrayLike,
        DtypeObj,
        IgnoreRaise,
    )

    from pandas.core.arrays import ExtensionArray

_dtype_obj = np.dtype(object)


@overload
def _astype_nansafe(
    arr: np.ndarray, dtype: np.dtype, copy: bool = ..., skipna: bool = ...
) -> np.ndarray:
    ...


@overload
def _astype_nansafe(
    arr: np.ndarray, dtype: ExtensionDtype, copy: bool = ..., skipna: bool = ...
) -> ExtensionArray:
    ...


def _astype_nansafe(
    arr: np.ndarray, dtype: DtypeObj, copy: bool = True, skipna: bool = False
) -> ArrayLike:
    """
    Cast the elements of an array to a given dtype a nan-safe manner.

    Parameters
    ----------
    arr : ndarray
    dtype : np.dtype or ExtensionDtype
    copy : bool, default True
        If False, a view will be attempted but may fail, if
        e.g. the item sizes don't align.
    skipna: bool, default False
        Whether or not we should skip NaN when casting as a string-type.

    Raises
    ------
    ValueError
        The dtype was a datetime64/timedelta64 dtype, but it had no unit.
    """

    # dispatch on extension dtype if needed
    if isinstance(dtype, ExtensionDtype):
        return dtype.construct_array_type()._from_sequence(arr, dtype=dtype, copy=copy)

    elif not isinstance(dtype, np.dtype):  # pragma: no cover
        raise ValueError("dtype must be np.dtype or ExtensionDtype")

    if arr.dtype.kind in "mM":
        from pandas.core.construction import ensure_wrapped_if_datetimelike

        arr = ensure_wrapped_if_datetimelike(arr)
        res = arr.astype(dtype, copy=copy)
        return np.asarray(res)

    if issubclass(dtype.type, str):
        shape = arr.shape
        if arr.ndim > 1:
            arr = arr.ravel()
        return lib.ensure_string_array(
            arr, skipna=skipna, convert_na_value=False
        ).reshape(shape)

    elif np.issubdtype(arr.dtype, np.floating) and dtype.kind in "iu":
        return _astype_float_to_int_nansafe(arr, dtype, copy)

    elif arr.dtype == object:
        # if we have a datetime/timedelta array of objects
        # then coerce to datetime64[ns] and use DatetimeArray.astype

        if lib.is_np_dtype(dtype, "M"):
            from pandas.core.arrays import DatetimeArray

            dta = DatetimeArray._from_sequence(arr, dtype=dtype)
            return dta._ndarray

        elif lib.is_np_dtype(dtype, "m"):
            from pandas.core.construction import ensure_wrapped_if_datetimelike

            # bc we know arr.dtype == object, this is equivalent to
            #  `np.asarray(to_timedelta(arr))`, but using a lower-level API that
            #  does not require a circular import.
            tdvals = array_to_timedelta64(arr).view("m8[ns]")

            tda = ensure_wrapped_if_datetimelike(tdvals)
            return tda.astype(dtype, copy=False)._ndarray

    if dtype.name in ("datetime64", "timedelta64"):
        msg = (
            f"The '{dtype.name}' dtype has no unit. Please pass in "
            f"'{dtype.name}[ns]' instead."
        )
        raise ValueError(msg)

    if copy or arr.dtype == object or dtype == object:
        # Explicit copy, or required since NumPy can't view from / to object.
        return arr.astype(dtype, copy=True)

    return arr.astype(dtype, copy=copy)


def _astype_float_to_int_nansafe(
    values: np.ndarray, dtype: np.dtype, copy: bool
) -> np.ndarray:
    """
    astype with a check preventing converting NaN to an meaningless integer value.
    """
    if not np.isfinite(values).all():
        raise IntCastingNaNError(
            "Cannot convert non-finite values (NA or inf) to integer"
        )
    if dtype.kind == "u":
        # GH#45151
        if not (values >= 0).all():
            raise ValueError(f"Cannot losslessly cast from {values.dtype} to {dtype}")
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=RuntimeWarning)
        return values.astype(dtype, copy=copy)


def astype_array(values: ArrayLike, dtype: DtypeObj, copy: bool = False) -> ArrayLike:
    """
    Cast array (ndarray or ExtensionArray) to the new dtype.

    Parameters
    ----------
    values : ndarray or ExtensionArray
    dtype : dtype object
    copy : bool, default False
        copy if indicated

    Returns
    -------
    ndarray or ExtensionArray
    """
    if values.dtype == dtype:
        if copy:
            return values.copy()
        return values

    if not isinstance(values, np.ndarray):
        # i.e. ExtensionArray
        values = values.astype(dtype, copy=copy)

    else:
        values = _astype_nansafe(values, dtype, copy=copy)

    # in pandas we don't store numpy str dtypes, so convert to object
    if isinstance(dtype, np.dtype) and issubclass(values.dtype.type, str):
        values = np.array(values, dtype=object)

    return values


def astype_array_safe(
    values: ArrayLike, dtype, copy: bool = False, errors: IgnoreRaise = "raise"
) -> ArrayLike:
    """
    Cast array (ndarray or ExtensionArray) to the new dtype.

    This basically is the implementation for DataFrame/Series.astype and
    includes all custom logic for pandas (NaN-safety, converting str to object,
    not allowing )

    Parameters
    ----------
    values : ndarray or ExtensionArray
    dtype : str, dtype convertible
    copy : bool, default False
        copy if indicated
    errors : str, {'raise', 'ignore'}, default 'raise'
        - ``raise`` : allow exceptions to be raised
        - ``ignore`` : suppress exceptions. On error return original object

    Returns
    -------
    ndarray or ExtensionArray
    """
    errors_legal_values = ("raise", "ignore")

    if errors not in errors_legal_values:
        invalid_arg = (
            "Expected value of kwarg 'errors' to be one of "
            f"{list(errors_legal_values)}. Supplied value is '{errors}'"
        )
        raise ValueError(invalid_arg)

    if inspect.isclass(dtype) and issubclass(dtype, ExtensionDtype):
        msg = (
            f"Expected an instance of {dtype.__name__}, "
            "but got the class instead. Try instantiating 'dtype'."
        )
        raise TypeError(msg)

    dtype = pandas_dtype(dtype)
    if isinstance(dtype, NumpyEADtype):
        # Ensure we don't end up with a NumpyExtensionArray
        dtype = dtype.numpy_dtype

    try:
        new_values = astype_array(values, dtype, copy=copy)
    except (ValueError, TypeError):
        # e.g. _astype_nansafe can fail on object-dtype of strings
        #  trying to convert to float
        if errors == "ignore":
            new_values = values
        else:
            raise

    return new_values


def astype_is_view(dtype: DtypeObj, new_dtype: DtypeObj) -> bool:
    """Checks if astype avoided copying the data.

    Parameters
    ----------
    dtype : Original dtype
    new_dtype : target dtype

    Returns
    -------
    True if new data is a view or not guaranteed to be a copy, False otherwise
    """
    if isinstance(dtype, np.dtype) and not isinstance(new_dtype, np.dtype):
        new_dtype, dtype = dtype, new_dtype

    if dtype == new_dtype:
        return True

    elif isinstance(dtype, np.dtype) and isinstance(new_dtype, np.dtype):
        # Only equal numpy dtypes avoid a copy
        return False

    elif is_string_dtype(dtype) and is_string_dtype(new_dtype):
        # Potentially! a view when converting from object to string
        return True

    elif is_object_dtype(dtype) and new_dtype.kind == "O":
        # When the underlying array has dtype object, we don't have to make a copy
        return True

    elif dtype.kind in "mM" and new_dtype.kind in "mM":
        dtype = getattr(dtype, "numpy_dtype", dtype)
        new_dtype = getattr(new_dtype, "numpy_dtype", new_dtype)
        return getattr(dtype, "unit", None) == getattr(new_dtype, "unit", None)

    numpy_dtype = getattr(dtype, "numpy_dtype", None)
    new_numpy_dtype = getattr(new_dtype, "numpy_dtype", None)

    if numpy_dtype is None and isinstance(dtype, np.dtype):
        numpy_dtype = dtype

    if new_numpy_dtype is None and isinstance(new_dtype, np.dtype):
        new_numpy_dtype = new_dtype

    if numpy_dtype is not None and new_numpy_dtype is not None:
        # if both have NumPy dtype or one of them is a numpy dtype
        # they are only a view when the numpy dtypes are equal, e.g.
        # int64 -> Int64 or int64[pyarrow]
        # int64 -> Int32 copies
        return numpy_dtype == new_numpy_dtype

    # Assume this is a view since we don't know for sure if a copy was made
    return True

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\metadata\pkg_resources.py ===
import email.message
import email.parser
import logging
import os
import zipfile
from typing import (
    Collection,
    Iterable,
    Iterator,
    List,
    Mapping,
    NamedTuple,
    Optional,
)

from pip._vendor import pkg_resources
from pip._vendor.packaging.requirements import Requirement
from pip._vendor.packaging.utils import NormalizedName, canonicalize_name
from pip._vendor.packaging.version import Version
from pip._vendor.packaging.version import parse as parse_version

from pip._internal.exceptions import InvalidWheel, NoneMetadataError, UnsupportedWheel
from pip._internal.utils.egg_link import egg_link_path_from_location
from pip._internal.utils.misc import display_path, normalize_path
from pip._internal.utils.wheel import parse_wheel, read_wheel_metadata_file

from .base import (
    BaseDistribution,
    BaseEntryPoint,
    BaseEnvironment,
    InfoPath,
    Wheel,
)

__all__ = ["NAME", "Distribution", "Environment"]

logger = logging.getLogger(__name__)

NAME = "pkg_resources"


class EntryPoint(NamedTuple):
    name: str
    value: str
    group: str


class InMemoryMetadata:
    """IMetadataProvider that reads metadata files from a dictionary.

    This also maps metadata decoding exceptions to our internal exception type.
    """

    def __init__(self, metadata: Mapping[str, bytes], wheel_name: str) -> None:
        self._metadata = metadata
        self._wheel_name = wheel_name

    def has_metadata(self, name: str) -> bool:
        return name in self._metadata

    def get_metadata(self, name: str) -> str:
        try:
            return self._metadata[name].decode()
        except UnicodeDecodeError as e:
            # Augment the default error with the origin of the file.
            raise UnsupportedWheel(
                f"Error decoding metadata for {self._wheel_name}: {e} in {name} file"
            )

    def get_metadata_lines(self, name: str) -> Iterable[str]:
        return pkg_resources.yield_lines(self.get_metadata(name))

    def metadata_isdir(self, name: str) -> bool:
        return False

    def metadata_listdir(self, name: str) -> List[str]:
        return []

    def run_script(self, script_name: str, namespace: str) -> None:
        pass


class Distribution(BaseDistribution):
    def __init__(self, dist: pkg_resources.Distribution) -> None:
        self._dist = dist
        # This is populated lazily, to avoid loading metadata for all possible
        # distributions eagerly.
        self.__extra_mapping: Optional[Mapping[NormalizedName, str]] = None

    @property
    def _extra_mapping(self) -> Mapping[NormalizedName, str]:
        if self.__extra_mapping is None:
            self.__extra_mapping = {
                canonicalize_name(extra): extra for extra in self._dist.extras
            }

        return self.__extra_mapping

    @classmethod
    def from_directory(cls, directory: str) -> BaseDistribution:
        dist_dir = directory.rstrip(os.sep)

        # Build a PathMetadata object, from path to metadata. :wink:
        base_dir, dist_dir_name = os.path.split(dist_dir)
        metadata = pkg_resources.PathMetadata(base_dir, dist_dir)

        # Determine the correct Distribution object type.
        if dist_dir.endswith(".egg-info"):
            dist_cls = pkg_resources.Distribution
            dist_name = os.path.splitext(dist_dir_name)[0]
        else:
            assert dist_dir.endswith(".dist-info")
            dist_cls = pkg_resources.DistInfoDistribution
            dist_name = os.path.splitext(dist_dir_name)[0].split("-")[0]

        dist = dist_cls(base_dir, project_name=dist_name, metadata=metadata)
        return cls(dist)

    @classmethod
    def from_metadata_file_contents(
        cls,
        metadata_contents: bytes,
        filename: str,
        project_name: str,
    ) -> BaseDistribution:
        metadata_dict = {
            "METADATA": metadata_contents,
        }
        dist = pkg_resources.DistInfoDistribution(
            location=filename,
            metadata=InMemoryMetadata(metadata_dict, filename),
            project_name=project_name,
        )
        return cls(dist)

    @classmethod
    def from_wheel(cls, wheel: Wheel, name: str) -> BaseDistribution:
        try:
            with wheel.as_zipfile() as zf:
                info_dir, _ = parse_wheel(zf, name)
                metadata_dict = {
                    path.split("/", 1)[-1]: read_wheel_metadata_file(zf, path)
                    for path in zf.namelist()
                    if path.startswith(f"{info_dir}/")
                }
        except zipfile.BadZipFile as e:
            raise InvalidWheel(wheel.location, name) from e
        except UnsupportedWheel as e:
            raise UnsupportedWheel(f"{name} has an invalid wheel, {e}")
        dist = pkg_resources.DistInfoDistribution(
            location=wheel.location,
            metadata=InMemoryMetadata(metadata_dict, wheel.location),
            project_name=name,
        )
        return cls(dist)

    @property
    def location(self) -> Optional[str]:
        return self._dist.location

    @property
    def installed_location(self) -> Optional[str]:
        egg_link = egg_link_path_from_location(self.raw_name)
        if egg_link:
            location = egg_link
        elif self.location:
            location = self.location
        else:
            return None
        return normalize_path(location)

    @property
    def info_location(self) -> Optional[str]:
        return self._dist.egg_info

    @property
    def installed_by_distutils(self) -> bool:
        # A distutils-installed distribution is provided by FileMetadata. This
        # provider has a "path" attribute not present anywhere else. Not the
        # best introspection logic, but pip has been doing this for a long time.
        try:
            return bool(self._dist._provider.path)
        except AttributeError:
            return False

    @property
    def canonical_name(self) -> NormalizedName:
        return canonicalize_name(self._dist.project_name)

    @property
    def version(self) -> Version:
        return parse_version(self._dist.version)

    @property
    def raw_version(self) -> str:
        return self._dist.version

    def is_file(self, path: InfoPath) -> bool:
        return self._dist.has_metadata(str(path))

    def iter_distutils_script_names(self) -> Iterator[str]:
        yield from self._dist.metadata_listdir("scripts")

    def read_text(self, path: InfoPath) -> str:
        name = str(path)
        if not self._dist.has_metadata(name):
            raise FileNotFoundError(name)
        content = self._dist.get_metadata(name)
        if content is None:
            raise NoneMetadataError(self, name)
        return content

    def iter_entry_points(self) -> Iterable[BaseEntryPoint]:
        for group, entries in self._dist.get_entry_map().items():
            for name, entry_point in entries.items():
                name, _, value = str(entry_point).partition("=")
                yield EntryPoint(name=name.strip(), value=value.strip(), group=group)

    def _metadata_impl(self) -> email.message.Message:
        """
        :raises NoneMetadataError: if the distribution reports `has_metadata()`
            True but `get_metadata()` returns None.
        """
        if isinstance(self._dist, pkg_resources.DistInfoDistribution):
            metadata_name = "METADATA"
        else:
            metadata_name = "PKG-INFO"
        try:
            metadata = self.read_text(metadata_name)
        except FileNotFoundError:
            if self.location:
                displaying_path = display_path(self.location)
            else:
                displaying_path = repr(self.location)
            logger.warning("No metadata found in %s", displaying_path)
            metadata = ""
        feed_parser = email.parser.FeedParser()
        feed_parser.feed(metadata)
        return feed_parser.close()

    def iter_dependencies(self, extras: Collection[str] = ()) -> Iterable[Requirement]:
        if extras:
            relevant_extras = set(self._extra_mapping) & set(
                map(canonicalize_name, extras)
            )
            extras = [self._extra_mapping[extra] for extra in relevant_extras]
        return self._dist.requires(extras)

    def iter_provided_extras(self) -> Iterable[NormalizedName]:
        return self._extra_mapping.keys()


class Environment(BaseEnvironment):
    def __init__(self, ws: pkg_resources.WorkingSet) -> None:
        self._ws = ws

    @classmethod
    def default(cls) -> BaseEnvironment:
        return cls(pkg_resources.working_set)

    @classmethod
    def from_paths(cls, paths: Optional[List[str]]) -> BaseEnvironment:
        return cls(pkg_resources.WorkingSet(paths))

    def _iter_distributions(self) -> Iterator[BaseDistribution]:
        for dist in self._ws:
            yield Distribution(dist)

    def _search_distribution(self, name: str) -> Optional[BaseDistribution]:
        """Find a distribution matching the ``name`` in the environment.

        This searches from *all* distributions available in the environment, to
        match the behavior of ``pkg_resources.get_distribution()``.
        """
        canonical_name = canonicalize_name(name)
        for dist in self.iter_all_distributions():
            if dist.canonical_name == canonical_name:
                return dist
        return None

    def get_distribution(self, name: str) -> Optional[BaseDistribution]:
        # Search the distribution by looking through the working set.
        dist = self._search_distribution(name)
        if dist:
            return dist

        # If distribution could not be found, call working_set.require to
        # update the working set, and try to find the distribution again.
        # This might happen for e.g. when you install a package twice, once
        # using setup.py develop and again using setup.py install. Now when
        # running pip uninstall twice, the package gets removed from the
        # working set in the first uninstall, so we have to populate the
        # working set again so that pip knows about it and the packages gets
        # picked up and is successfully uninstalled the second time too.
        try:
            # We didn't pass in any version specifiers, so this can never
            # raise pkg_resources.VersionConflict.
            self._ws.require(name)
        except pkg_resources.DistributionNotFound:
            return None
        return self._search_distribution(name)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\combinatorics\generators.py ===
from sympy.combinatorics.permutations import Permutation
from sympy.core.symbol import symbols
from sympy.matrices import Matrix
from sympy.utilities.iterables import variations, rotate_left


def symmetric(n):
    """
    Generates the symmetric group of order n, Sn.

    Examples
    ========

    >>> from sympy.combinatorics.generators import symmetric
    >>> list(symmetric(3))
    [(2), (1 2), (2)(0 1), (0 1 2), (0 2 1), (0 2)]
    """
    yield from (Permutation(perm) for perm in variations(range(n), n))


def cyclic(n):
    """
    Generates the cyclic group of order n, Cn.

    Examples
    ========

    >>> from sympy.combinatorics.generators import cyclic
    >>> list(cyclic(5))
    [(4), (0 1 2 3 4), (0 2 4 1 3),
     (0 3 1 4 2), (0 4 3 2 1)]

    See Also
    ========

    dihedral
    """
    gen = list(range(n))
    for i in range(n):
        yield Permutation(gen)
        gen = rotate_left(gen, 1)


def alternating(n):
    """
    Generates the alternating group of order n, An.

    Examples
    ========

    >>> from sympy.combinatorics.generators import alternating
    >>> list(alternating(3))
    [(2), (0 1 2), (0 2 1)]
    """
    for perm in variations(range(n), n):
        p = Permutation(perm)
        if p.is_even:
            yield p


def dihedral(n):
    """
    Generates the dihedral group of order 2n, Dn.

    The result is given as a subgroup of Sn, except for the special cases n=1
    (the group S2) and n=2 (the Klein 4-group) where that's not possible
    and embeddings in S2 and S4 respectively are given.

    Examples
    ========

    >>> from sympy.combinatorics.generators import dihedral
    >>> list(dihedral(3))
    [(2), (0 2), (0 1 2), (1 2), (0 2 1), (2)(0 1)]

    See Also
    ========

    cyclic
    """
    if n == 1:
        yield Permutation([0, 1])
        yield Permutation([1, 0])
    elif n == 2:
        yield Permutation([0, 1, 2, 3])
        yield Permutation([1, 0, 3, 2])
        yield Permutation([2, 3, 0, 1])
        yield Permutation([3, 2, 1, 0])
    else:
        gen = list(range(n))
        for i in range(n):
            yield Permutation(gen)
            yield Permutation(gen[::-1])
            gen = rotate_left(gen, 1)


def rubik_cube_generators():
    """Return the permutations of the 3x3 Rubik's cube, see
    https://www.gap-system.org/Doc/Examples/rubik.html
    """
    a = [
        [(1, 3, 8, 6), (2, 5, 7, 4), (9, 33, 25, 17), (10, 34, 26, 18),
         (11, 35, 27, 19)],
        [(9, 11, 16, 14), (10, 13, 15, 12), (1, 17, 41, 40), (4, 20, 44, 37),
         (6, 22, 46, 35)],
        [(17, 19, 24, 22), (18, 21, 23, 20), (6, 25, 43, 16), (7, 28, 42, 13),
         (8, 30, 41, 11)],
        [(25, 27, 32, 30), (26, 29, 31, 28), (3, 38, 43, 19), (5, 36, 45, 21),
         (8, 33, 48, 24)],
        [(33, 35, 40, 38), (34, 37, 39, 36), (3, 9, 46, 32), (2, 12, 47, 29),
         (1, 14, 48, 27)],
        [(41, 43, 48, 46), (42, 45, 47, 44), (14, 22, 30, 38),
         (15, 23, 31, 39), (16, 24, 32, 40)]
    ]
    return [Permutation([[i - 1 for i in xi] for xi in x], size=48) for x in a]


def rubik(n):
    """Return permutations for an nxn Rubik's cube.

    Permutations returned are for rotation of each of the slice
    from the face up to the last face for each of the 3 sides (in this order):
    front, right and bottom. Hence, the first n - 1 permutations are for the
    slices from the front.
    """

    if n < 2:
        raise ValueError('dimension of cube must be > 1')

    # 1-based reference to rows and columns in Matrix
    def getr(f, i):
        return faces[f].col(n - i)

    def getl(f, i):
        return faces[f].col(i - 1)

    def getu(f, i):
        return faces[f].row(i - 1)

    def getd(f, i):
        return faces[f].row(n - i)

    def setr(f, i, s):
        faces[f][:, n - i] = Matrix(n, 1, s)

    def setl(f, i, s):
        faces[f][:, i - 1] = Matrix(n, 1, s)

    def setu(f, i, s):
        faces[f][i - 1, :] = Matrix(1, n, s)

    def setd(f, i, s):
        faces[f][n - i, :] = Matrix(1, n, s)

    # motion of a single face
    def cw(F, r=1):
        for _ in range(r):
            face = faces[F]
            rv = []
            for c in range(n):
                for r in range(n - 1, -1, -1):
                    rv.append(face[r, c])
            faces[F] = Matrix(n, n, rv)

    def ccw(F):
        cw(F, 3)

    # motion of plane i from the F side;
    # fcw(0) moves the F face, fcw(1) moves the plane
    # just behind the front face, etc...
    def fcw(i, r=1):
        for _ in range(r):
            if i == 0:
                cw(F)
            i += 1
            temp = getr(L, i)
            setr(L, i, list(getu(D, i)))
            setu(D, i, list(reversed(getl(R, i))))
            setl(R, i, list(getd(U, i)))
            setd(U, i, list(reversed(temp)))
            i -= 1

    def fccw(i):
        fcw(i, 3)

    # motion of the entire cube from the F side
    def FCW(r=1):
        for _ in range(r):
            cw(F)
            ccw(B)
            cw(U)
            t = faces[U]
            cw(L)
            faces[U] = faces[L]
            cw(D)
            faces[L] = faces[D]
            cw(R)
            faces[D] = faces[R]
            faces[R] = t

    def FCCW():
        FCW(3)

    # motion of the entire cube from the U side
    def UCW(r=1):
        for _ in range(r):
            cw(U)
            ccw(D)
            t = faces[F]
            faces[F] = faces[R]
            faces[R] = faces[B]
            faces[B] = faces[L]
            faces[L] = t

    def UCCW():
        UCW(3)

    # defining the permutations for the cube

    U, F, R, B, L, D = names = symbols('U, F, R, B, L, D')

    # the faces are represented by nxn matrices
    faces = {}
    count = 0
    for fi in range(6):
        f = []
        for a in range(n**2):
            f.append(count)
            count += 1
        faces[names[fi]] = Matrix(n, n, f)

    # this will either return the value of the current permutation
    # (show != 1) or else append the permutation to the group, g
    def perm(show=0):
        # add perm to the list of perms
        p = []
        for f in names:
            p.extend(faces[f])
        if show:
            return p
        g.append(Permutation(p))

    g = []  # container for the group's permutations
    I = list(range(6*n**2))  # the identity permutation used for checking

    # define permutations corresponding to cw rotations of the planes
    # up TO the last plane from that direction; by not including the
    # last plane, the orientation of the cube is maintained.

    # F slices
    for i in range(n - 1):
        fcw(i)
        perm()
        fccw(i)  # restore
    assert perm(1) == I

    # R slices
    # bring R to front
    UCW()
    for i in range(n - 1):
        fcw(i)
        # put it back in place
        UCCW()
        # record
        perm()
        # restore
        # bring face to front
        UCW()
        fccw(i)
    # restore
    UCCW()
    assert perm(1) == I

    # D slices
    # bring up bottom
    FCW()
    UCCW()
    FCCW()
    for i in range(n - 1):
        # turn strip
        fcw(i)
        # put bottom back on the bottom
        FCW()
        UCW()
        FCCW()
        # record
        perm()
        # restore
        # bring up bottom
        FCW()
        UCCW()
        FCCW()
        # turn strip
        fccw(i)
    # put bottom back on the bottom
    FCW()
    UCW()
    FCCW()
    assert perm(1) == I

    return g

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\utilities\_compilation\runners.py ===
from __future__ import annotations
from typing import Callable, Optional

from collections import OrderedDict
import os
import re
import subprocess
import warnings

from .util import (
    find_binary_of_command, unique_list, CompileError
)


class CompilerRunner:
    """ CompilerRunner base class.

    Parameters
    ==========

    sources : list of str
        Paths to sources.
    out : str
    flags : iterable of str
        Compiler flags.
    run_linker : bool
    compiler_name_exe : (str, str) tuple
        Tuple of compiler name &  command to call.
    cwd : str
        Path of root of relative paths.
    include_dirs : list of str
        Include directories.
    libraries : list of str
        Libraries to link against.
    library_dirs : list of str
        Paths to search for shared libraries.
    std : str
        Standard string, e.g. ``'c++11'``, ``'c99'``, ``'f2003'``.
    define: iterable of strings
        macros to define
    undef : iterable of strings
        macros to undefine
    preferred_vendor : string
        name of preferred vendor e.g. 'gnu' or 'intel'

    Methods
    =======

    run():
        Invoke compilation as a subprocess.

    """

    environ_key_compiler: str  # e.g. 'CC', 'CXX', ...
    environ_key_flags: str  # e.g. 'CFLAGS', 'CXXFLAGS', ...
    environ_key_ldflags: str = "LDFLAGS"  # typically 'LDFLAGS'

    # Subclass to vendor/binary dict
    compiler_dict: dict[str, str]

    # Standards should be a tuple of supported standards
    # (first one will be the default)
    standards: tuple[None | str, ...]

    # Subclass to dict of binary/formater-callback
    std_formater: dict[str, Callable[[Optional[str]], str]]

    # subclass to be e.g. {'gcc': 'gnu', ...}
    compiler_name_vendor_mapping: dict[str, str]

    def __init__(self, sources, out, flags=None, run_linker=True, compiler=None, cwd='.',
                 include_dirs=None, libraries=None, library_dirs=None, std=None, define=None,
                 undef=None, strict_aliasing=None, preferred_vendor=None, linkline=None, **kwargs):
        if isinstance(sources, str):
            raise ValueError("Expected argument sources to be a list of strings.")
        self.sources = list(sources)
        self.out = out
        self.flags = flags or []
        if os.environ.get(self.environ_key_flags):
            self.flags += os.environ[self.environ_key_flags].split()
        self.cwd = cwd
        if compiler:
            self.compiler_name, self.compiler_binary = compiler
        elif os.environ.get(self.environ_key_compiler):
            self.compiler_binary = os.environ[self.environ_key_compiler]
            for k, v in self.compiler_dict.items():
                if k in self.compiler_binary:
                    self.compiler_vendor = k
                    self.compiler_name = v
                    break
            else:
                self.compiler_vendor, self.compiler_name = list(self.compiler_dict.items())[0]
                warnings.warn("failed to determine what kind of compiler %s is, assuming %s" %
                              (self.compiler_binary, self.compiler_name))
        else:
            # Find a compiler
            if preferred_vendor is None:
                preferred_vendor = os.environ.get('SYMPY_COMPILER_VENDOR', None)
            self.compiler_name, self.compiler_binary, self.compiler_vendor = self.find_compiler(preferred_vendor)
            if self.compiler_binary is None:
                raise ValueError("No compiler found (searched: {})".format(', '.join(self.compiler_dict.values())))
        self.define = define or []
        self.undef = undef or []
        self.include_dirs = include_dirs or []
        self.libraries = libraries or []
        self.library_dirs = library_dirs or []
        self.std = std or self.standards[0]
        self.run_linker = run_linker
        if self.run_linker:
            # both gnu and intel compilers use '-c' for disabling linker
            self.flags = list(filter(lambda x: x != '-c', self.flags))
        else:
            if '-c' not in self.flags:
                self.flags.append('-c')

        if self.std:
            self.flags.append(self.std_formater[
                self.compiler_name](self.std))

        self.linkline = (linkline or []) + [lf for lf in map(
            str.strip, os.environ.get(self.environ_key_ldflags, "").split()
        ) if lf != ""]

        if strict_aliasing is not None:
            nsa_re = re.compile("no-strict-aliasing$")
            sa_re = re.compile("strict-aliasing$")
            if strict_aliasing is True:
                if any(map(nsa_re.match, flags)):
                    raise CompileError("Strict aliasing cannot be both enforced and disabled")
                elif any(map(sa_re.match, flags)):
                    pass  # already enforced
                else:
                    flags.append('-fstrict-aliasing')
            elif strict_aliasing is False:
                if any(map(nsa_re.match, flags)):
                    pass  # already disabled
                else:
                    if any(map(sa_re.match, flags)):
                        raise CompileError("Strict aliasing cannot be both enforced and disabled")
                    else:
                        flags.append('-fno-strict-aliasing')
            else:
                msg = "Expected argument strict_aliasing to be True/False, got {}"
                raise ValueError(msg.format(strict_aliasing))

    @classmethod
    def find_compiler(cls, preferred_vendor=None):
        """ Identify a suitable C/fortran/other compiler. """
        candidates = list(cls.compiler_dict.keys())
        if preferred_vendor:
            if preferred_vendor in candidates:
                candidates = [preferred_vendor]+candidates
            else:
                raise ValueError("Unknown vendor {}".format(preferred_vendor))
        name, path = find_binary_of_command([cls.compiler_dict[x] for x in candidates])
        return name, path, cls.compiler_name_vendor_mapping[name]

    def cmd(self):
        """ List of arguments (str) to be passed to e.g. ``subprocess.Popen``. """
        cmd = (
            [self.compiler_binary] +
            self.flags +
            ['-U'+x for x in self.undef] +
            ['-D'+x for x in self.define] +
            ['-I'+x for x in self.include_dirs] +
            self.sources
        )
        if self.run_linker:
            cmd += (['-L'+x for x in self.library_dirs] +
                    ['-l'+x for x in self.libraries] +
                    self.linkline)
        counted = []
        for envvar in re.findall(r'\$\{(\w+)\}', ' '.join(cmd)):
            if os.getenv(envvar) is None:
                if envvar not in counted:
                    counted.append(envvar)
                    msg = "Environment variable '{}' undefined.".format(envvar)
                    raise CompileError(msg)
        return cmd

    def run(self):
        self.flags = unique_list(self.flags)

        # Append output flag and name to tail of flags
        self.flags.extend(['-o', self.out])
        env = os.environ.copy()
        env['PWD'] = self.cwd

        # NOTE: intel compilers seems to need shell=True
        p = subprocess.Popen(' '.join(self.cmd()),
                             shell=True,
                             cwd=self.cwd,
                             stdin=subprocess.PIPE,
                             stdout=subprocess.PIPE,
                             stderr=subprocess.STDOUT,
                             env=env)
        comm = p.communicate()
        try:
            self.cmd_outerr = comm[0].decode('utf-8')
        except UnicodeDecodeError:
            self.cmd_outerr = comm[0].decode('iso-8859-1')  # win32
        self.cmd_returncode = p.returncode

        # Error handling
        if self.cmd_returncode != 0:
            msg = "Error executing '{}' in {} (exited status {}):\n {}\n".format(
                ' '.join(self.cmd()), self.cwd, str(self.cmd_returncode), self.cmd_outerr
            )
            raise CompileError(msg)

        return self.cmd_outerr, self.cmd_returncode


class CCompilerRunner(CompilerRunner):

    environ_key_compiler = 'CC'
    environ_key_flags = 'CFLAGS'

    compiler_dict = OrderedDict([
        ('gnu', 'gcc'),
        ('intel', 'icc'),
        ('llvm', 'clang'),
    ])

    standards = ('c89', 'c90', 'c99', 'c11')  # First is default

    std_formater = {
        'gcc': '-std={}'.format,
        'icc': '-std={}'.format,
        'clang': '-std={}'.format,
    }

    compiler_name_vendor_mapping = {
        'gcc': 'gnu',
        'icc': 'intel',
        'clang': 'llvm'
    }


def _mk_flag_filter(cmplr_name):  # helper for class initialization
    not_welcome = {'g++': ("Wimplicit-interface",)}  # "Wstrict-prototypes",)}
    if cmplr_name in not_welcome:
        def fltr(x):
            for nw in not_welcome[cmplr_name]:
                if nw in x:
                    return False
            return True
    else:
        def fltr(x):
            return True
    return fltr


class CppCompilerRunner(CompilerRunner):

    environ_key_compiler = 'CXX'
    environ_key_flags = 'CXXFLAGS'

    compiler_dict = OrderedDict([
        ('gnu', 'g++'),
        ('intel', 'icpc'),
        ('llvm', 'clang++'),
    ])

    # First is the default, c++0x == c++11
    standards = ('c++98', 'c++0x')

    std_formater = {
        'g++': '-std={}'.format,
        'icpc': '-std={}'.format,
        'clang++': '-std={}'.format,
    }

    compiler_name_vendor_mapping = {
        'g++': 'gnu',
        'icpc': 'intel',
        'clang++': 'llvm'
    }


class FortranCompilerRunner(CompilerRunner):

    environ_key_compiler = 'FC'
    environ_key_flags = 'FFLAGS'

    standards = (None, 'f77', 'f95', 'f2003', 'f2008')

    std_formater = {
        'gfortran': lambda x: '-std=gnu' if x is None else '-std=legacy' if x == 'f77' else '-std={}'.format(x),
        'ifort': lambda x: '-stand f08' if x is None else '-stand f{}'.format(x[-2:]),  # f2008 => f08
    }

    compiler_dict = OrderedDict([
        ('gnu', 'gfortran'),
        ('intel', 'ifort'),
    ])

    compiler_name_vendor_mapping = {
        'gfortran': 'gnu',
        'ifort': 'intel',
    }

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_attention_vae.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
from logging import getLogger

import numpy as np
from fusion_base import Fusion
from onnx import NodeProto, TensorProto, helper, numpy_helper
from onnx_model import OnnxModel

logger = getLogger(__name__)


class FusionAttentionVae(Fusion):
    """
    Fuse Attention subgraph of Vae Decoder into one Attention node.
    """

    def __init__(self, model: OnnxModel, hidden_size: int, num_heads: int):
        super().__init__(model, "Attention", ["Softmax"])
        self.hidden_size = hidden_size
        self.num_heads = num_heads

        # Flags to show warning only once
        self.num_heads_warning = True
        self.hidden_size_warning = True

    def get_num_heads_and_hidden_size(self, reshape_q: NodeProto, add_q: NodeProto) -> tuple[int, int]:
        """Detect num_heads and hidden_size from a reshape node.

        Args:
            reshape_q (NodeProto): reshape node for Q
            add_q (NodeProto): add node for Q

        Returns:
            Tuple[int, int]: num_heads and hidden_size
        """
        concat = self.model.get_parent(reshape_q, 1)
        if concat is None or len(concat.input) != 4:
            return self.num_heads, self.hidden_size  # Fall back to user specified value

        value = self.model.get_constant_value(concat.input[2])
        if not (value is not None and isinstance(value, np.ndarray) and value.size == 1):
            return self.num_heads, self.hidden_size  # Fall back to user specified value
        num_heads = int(value)
        if num_heads <= 0:
            return self.num_heads, self.hidden_size  # Fall back to user specified value

        _, bias = self.model.get_constant_input(add_q)
        if (bias is None) or (not isinstance(bias, np.ndarray)) or bias.ndim != 1:
            return self.num_heads, self.hidden_size  # Fall back to user specified value

        hidden_size = bias.shape[0]

        if self.num_heads > 0 and num_heads != self.num_heads:
            if self.num_heads_warning:
                logger.warning(
                    "Detected number of attention heads is %d. Ignore --num_heads %d", num_heads, self.num_heads
                )
                self.num_heads_warning = False  # Do not show the warning more than once

        if self.hidden_size > 0 and hidden_size != self.hidden_size:
            if self.hidden_size_warning:
                logger.warning("Detected hidden size is %d. Ignore --hidden_size %d", hidden_size, self.hidden_size)
                self.hidden_size_warning = False  # Do not show the warning more than once

        return num_heads, hidden_size

    def create_attention_node(
        self,
        q_matmul: NodeProto,
        q_add: NodeProto,
        k_matmul: NodeProto,
        k_add: NodeProto,
        v_matmul: NodeProto,
        v_add: NodeProto,
        num_heads: int,
        hidden_size: int,
        input_name: str,
        output_name: str,
    ) -> NodeProto | None:
        """Create an Attention node.

        Args:
            q_matmul (NodeProto): MatMul node in fully connection for Q
            q_add (NodeProto): Add bias node in fully connection for Q
            k_matmul (NodeProto): MatMul node in fully connection for K
            k_add (NodeProto): Add bias node in fully connection for K
            v_matmul (NodeProto): MatMul node in fully connection for V
            v_add (NodeProto): Add bias node in fully connection for V
            num_heads (int): number of attention heads. If a model is pruned, it is the number of heads after pruning.
            hidden_size (int): hidden dimension. If a model is pruned, it is the hidden dimension after pruning.
            input_name (str): input name
            output_name (str): output name

        Returns:
            Union[NodeProto, None]: the node created or None if failed.
        """
        if q_matmul.input[0] != input_name or k_matmul.input[0] != input_name or v_matmul.input[0] != input_name:
            logger.debug(
                "For self attention, input hidden state for q and k/v shall be same. Got %s, %s, %s",
                q_matmul.input[0],
                k_matmul.input[0],
                v_matmul.input[0],
            )
            return None

        if hidden_size > 0 and (hidden_size % num_heads) != 0:
            logger.debug("input hidden size %d is not a multiple of num of heads %d", hidden_size, num_heads)
            return None

        q_weight_tensor = self.model.get_initializer(q_matmul.input[1])
        k_weight_tensor = self.model.get_initializer(k_matmul.input[1])
        v_weight_tensor = self.model.get_initializer(v_matmul.input[1])
        if not (q_weight_tensor and k_weight_tensor and v_weight_tensor):
            return None

        q_bias_tensor = self.model.get_initializer(q_add.input[1]) or self.model.get_initializer(q_add.input[0])
        k_bias_tensor = self.model.get_initializer(k_add.input[1]) or self.model.get_initializer(k_add.input[0])
        v_bias_tensor = self.model.get_initializer(v_add.input[1]) or self.model.get_initializer(v_add.input[0])

        q_bias = numpy_helper.to_array(q_bias_tensor)
        k_bias = numpy_helper.to_array(k_bias_tensor)
        v_bias = numpy_helper.to_array(v_bias_tensor)

        q_bias_shape = np.prod(q_bias.shape)
        k_bias_shape = np.prod(k_bias.shape)
        v_bias_shape = np.prod(v_bias.shape)

        # Sometimes weights are stored in fp16
        if q_weight_tensor.data_type == 10:
            logger.debug("weights are in fp16. Please run fp16 conversion after optimization")
            return None

        q_weight = numpy_helper.to_array(q_weight_tensor)
        k_weight = numpy_helper.to_array(k_weight_tensor)
        v_weight = numpy_helper.to_array(v_weight_tensor)

        # assert q and k have same shape as expected
        if q_weight.shape != k_weight.shape or q_weight.shape != v_weight.shape:
            return None

        qw_in_size = q_weight.shape[0]
        kw_in_size = k_weight.shape[0]
        vw_in_size = v_weight.shape[0]

        assert qw_in_size == kw_in_size and kw_in_size == vw_in_size

        if hidden_size > 0 and hidden_size != qw_in_size:
            raise ValueError(
                f"Input hidden size ({hidden_size}) is not same as weight dimension of q,k,v ({qw_in_size}). "
                "Please provide a correct input hidden size or pass in 0"
            )

        # All the matrices can have the same shape or q, k matrics can have the same shape with v being different
        # For 2d weights, the shapes would be [in_size, out_size].
        # For 3d weights, shape would be [in_size, a, b] where a*b = out_size
        qw_out_size = np.prod(q_weight.shape[1:])

        qkv_weight = np.stack((q_weight, k_weight, v_weight), axis=1)
        qkv_weight_dim = 3 * int(qw_out_size)

        attention_node_name = self.model.create_node_name("Attention")

        assert q_bias_shape == k_bias_shape == v_bias_shape

        qkv_bias_dim = 0
        qkv_bias = np.stack((q_bias, k_bias, v_bias), axis=0)
        qkv_bias_dim = 3 * q_bias_shape

        self.add_initializer(
            name=attention_node_name + "_qkv_weight",
            data_type=TensorProto.FLOAT,
            dims=[qw_in_size, qkv_weight_dim],
            vals=qkv_weight,
        )

        # No bias, use zeros
        qkv_bias = np.zeros([3, hidden_size], dtype=np.float32)
        qkv_bias_dim = 3 * hidden_size

        self.add_initializer(
            name=attention_node_name + "_qkv_bias",
            data_type=TensorProto.FLOAT,
            dims=[qkv_bias_dim],
            vals=qkv_bias,
        )

        attention_inputs = [
            input_name,
            attention_node_name + "_qkv_weight",
            attention_node_name + "_qkv_bias",
        ]

        attention_node = helper.make_node(
            "Attention",
            inputs=attention_inputs,
            outputs=[output_name],
            name=attention_node_name,
        )
        attention_node.domain = "com.microsoft"
        attention_node.attribute.extend([helper.make_attribute("num_heads", num_heads)])

        self.increase_counter("Attention (self attention)")
        return attention_node

    def fuse(self, softmax_node, input_name_to_nodes, output_name_to_node):
        matmul_qkv = self.model.find_first_child_by_type(softmax_node, "MatMul", input_name_to_nodes, recursive=False)
        if matmul_qkv is None:
            return

        reshape_qkv = self.model.find_first_child_by_type(matmul_qkv, "Reshape", input_name_to_nodes, recursive=False)
        if reshape_qkv is None:
            return

        transpose_qkv = self.model.find_first_child_by_type(
            reshape_qkv, "Transpose", input_name_to_nodes, recursive=False
        )
        if transpose_qkv is None:
            return

        reshape_out = self.model.find_first_child_by_type(
            transpose_qkv, "Reshape", input_name_to_nodes, recursive=False
        )
        if reshape_out is None:
            return

        matmul_out = self.model.find_first_child_by_type(reshape_out, "MatMul", input_name_to_nodes, recursive=False)
        if matmul_out is None:
            return

        add_out = self.model.find_first_child_by_type(matmul_out, "Add", input_name_to_nodes, recursive=False)
        if add_out is None:
            return

        transpose_out = self.model.find_first_child_by_type(add_out, "Transpose", input_name_to_nodes, recursive=False)
        if transpose_out is None:
            return

        v_nodes = self.model.match_parent_path(
            matmul_qkv, ["Reshape", "Transpose", "Reshape", "Add", "MatMul"], [1, 0, 0, 0, None]
        )
        if v_nodes is None:
            logger.debug("fuse_attention: failed to match v path")
            return
        (_, _, _, add_v, matmul_v) = v_nodes

        qk_nodes = self.model.match_parent_path(matmul_qkv, ["Softmax", "Add", "Mul", "MatMul"], [0, 0, 0, 0])
        if qk_nodes is not None:
            (_softmax_qk, _add_zero, _mul_qk, matmul_qk) = qk_nodes
        else:
            logger.debug("fuse_attention: failed to match qk path")
            return

        q_nodes = self.model.match_parent_path(
            matmul_qk, ["Reshape", "Transpose", "Reshape", "Add", "MatMul"], [0, 0, 0, 0, None]
        )
        if q_nodes is None:
            logger.debug("fuse_attention: failed to match q path")
            return
        (_, _transpose_q, reshape_q, add_q, matmul_q) = q_nodes
        k_nodes = self.model.match_parent_path(
            matmul_qk, ["Transpose", "Reshape", "Transpose", "Reshape", "Add", "MatMul"], [1, 0, 0, 0, 0, None]
        )
        if k_nodes is None:
            logger.debug("fuse_attention: failed to match k path")
            return
        (_, _, _, _, add_k, matmul_k) = k_nodes

        attention_last_node = reshape_out

        q_num_heads, q_hidden_size = self.get_num_heads_and_hidden_size(reshape_q, add_q)
        if q_num_heads <= 0:
            logger.debug("fuse_attention: failed to detect num_heads")
            return

        # number of heads are same for all the paths, hence to create attention node, we pass the q_num_heads
        new_node = self.create_attention_node(
            matmul_q,
            add_q,
            matmul_k,
            add_k,
            matmul_v,
            add_v,
            q_num_heads,
            q_hidden_size,
            matmul_q.input[0],
            attention_last_node.output[0],
        )
        if new_node is None:
            return

        self.nodes_to_add.append(new_node)
        self.node_name_to_graph_name[new_node.name] = self.this_graph_name

        self.nodes_to_remove.extend([attention_last_node, transpose_qkv])

        # Use prune graph to remove nodes since they are shared by all attention nodes.
        self.prune_graph = True

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\orm\dynamic.py ===
# orm/dynamic.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php


"""Dynamic collection API.

Dynamic collections act like Query() objects for read operations and support
basic add/delete mutation.

.. legacy:: the "dynamic" loader is a legacy feature, superseded by the
 "write_only" loader.


"""

from __future__ import annotations

from typing import Any
from typing import Iterable
from typing import Iterator
from typing import List
from typing import Optional
from typing import Tuple
from typing import Type
from typing import TYPE_CHECKING
from typing import TypeVar
from typing import Union

from . import attributes
from . import exc as orm_exc
from . import relationships
from . import util as orm_util
from .base import PassiveFlag
from .query import Query
from .session import object_session
from .writeonly import AbstractCollectionWriter
from .writeonly import WriteOnlyAttributeImpl
from .writeonly import WriteOnlyHistory
from .writeonly import WriteOnlyLoader
from .. import util
from ..engine import result


if TYPE_CHECKING:
    from . import QueryableAttribute
    from .mapper import Mapper
    from .relationships import _RelationshipOrderByArg
    from .session import Session
    from .state import InstanceState
    from .util import AliasedClass
    from ..event import _Dispatch
    from ..sql.elements import ColumnElement

_T = TypeVar("_T", bound=Any)


class DynamicCollectionHistory(WriteOnlyHistory[_T]):
    def __init__(
        self,
        attr: DynamicAttributeImpl,
        state: InstanceState[_T],
        passive: PassiveFlag,
        apply_to: Optional[DynamicCollectionHistory[_T]] = None,
    ) -> None:
        if apply_to:
            coll = AppenderQuery(attr, state).autoflush(False)
            self.unchanged_items = util.OrderedIdentitySet(coll)
            self.added_items = apply_to.added_items
            self.deleted_items = apply_to.deleted_items
            self._reconcile_collection = True
        else:
            self.deleted_items = util.OrderedIdentitySet()
            self.added_items = util.OrderedIdentitySet()
            self.unchanged_items = util.OrderedIdentitySet()
            self._reconcile_collection = False


class DynamicAttributeImpl(WriteOnlyAttributeImpl):
    _supports_dynamic_iteration = True
    collection_history_cls = DynamicCollectionHistory[Any]
    query_class: Type[AppenderMixin[Any]]  # type: ignore[assignment]

    def __init__(
        self,
        class_: Union[Type[Any], AliasedClass[Any]],
        key: str,
        dispatch: _Dispatch[QueryableAttribute[Any]],
        target_mapper: Mapper[_T],
        order_by: _RelationshipOrderByArg,
        query_class: Optional[Type[AppenderMixin[_T]]] = None,
        **kw: Any,
    ) -> None:
        attributes.AttributeImpl.__init__(
            self, class_, key, None, dispatch, **kw
        )
        self.target_mapper = target_mapper
        if order_by:
            self.order_by = tuple(order_by)
        if not query_class:
            self.query_class = AppenderQuery
        elif AppenderMixin in query_class.mro():
            self.query_class = query_class
        else:
            self.query_class = mixin_user_query(query_class)


@relationships.RelationshipProperty.strategy_for(lazy="dynamic")
class DynaLoader(WriteOnlyLoader):
    impl_class = DynamicAttributeImpl


class AppenderMixin(AbstractCollectionWriter[_T]):
    """A mixin that expects to be mixing in a Query class with
    AbstractAppender.


    """

    query_class: Optional[Type[Query[_T]]] = None
    _order_by_clauses: Tuple[ColumnElement[Any], ...]

    def __init__(
        self, attr: DynamicAttributeImpl, state: InstanceState[_T]
    ) -> None:
        Query.__init__(
            self,  # type: ignore[arg-type]
            attr.target_mapper,
            None,
        )
        super().__init__(attr, state)

    @property
    def session(self) -> Optional[Session]:
        sess = object_session(self.instance)
        if sess is not None and sess.autoflush and self.instance in sess:
            sess.flush()
        if not orm_util.has_identity(self.instance):
            return None
        else:
            return sess

    @session.setter
    def session(self, session: Session) -> None:
        self.sess = session

    def _iter(self) -> Union[result.ScalarResult[_T], result.Result[_T]]:
        sess = self.session
        if sess is None:
            state = attributes.instance_state(self.instance)
            if state.detached:
                util.warn(
                    "Instance %s is detached, dynamic relationship cannot "
                    "return a correct result.   This warning will become "
                    "a DetachedInstanceError in a future release."
                    % (orm_util.state_str(state))
                )

            return result.IteratorResult(
                result.SimpleResultMetaData([self.attr.class_.__name__]),
                iter(
                    self.attr._get_collection_history(
                        attributes.instance_state(self.instance),
                        PassiveFlag.PASSIVE_NO_INITIALIZE,
                    ).added_items
                ),
                _source_supports_scalars=True,
            ).scalars()
        else:
            return self._generate(sess)._iter()

    if TYPE_CHECKING:

        def __iter__(self) -> Iterator[_T]: ...

    def __getitem__(self, index: Any) -> Union[_T, List[_T]]:
        sess = self.session
        if sess is None:
            return self.attr._get_collection_history(
                attributes.instance_state(self.instance),
                PassiveFlag.PASSIVE_NO_INITIALIZE,
            ).indexed(index)
        else:
            return self._generate(sess).__getitem__(index)  # type: ignore[no-any-return] # noqa: E501

    def count(self) -> int:
        sess = self.session
        if sess is None:
            return len(
                self.attr._get_collection_history(
                    attributes.instance_state(self.instance),
                    PassiveFlag.PASSIVE_NO_INITIALIZE,
                ).added_items
            )
        else:
            return self._generate(sess).count()

    def _generate(
        self,
        sess: Optional[Session] = None,
    ) -> Query[_T]:
        # note we're returning an entirely new Query class instance
        # here without any assignment capabilities; the class of this
        # query is determined by the session.
        instance = self.instance
        if sess is None:
            sess = object_session(instance)
            if sess is None:
                raise orm_exc.DetachedInstanceError(
                    "Parent instance %s is not bound to a Session, and no "
                    "contextual session is established; lazy load operation "
                    "of attribute '%s' cannot proceed"
                    % (orm_util.instance_str(instance), self.attr.key)
                )

        if self.query_class:
            query = self.query_class(self.attr.target_mapper, session=sess)
        else:
            query = sess.query(self.attr.target_mapper)

        query._where_criteria = self._where_criteria
        query._from_obj = self._from_obj
        query._order_by_clauses = self._order_by_clauses

        return query

    def add_all(self, iterator: Iterable[_T]) -> None:
        """Add an iterable of items to this :class:`_orm.AppenderQuery`.

        The given items will be persisted to the database in terms of
        the parent instance's collection on the next flush.

        This method is provided to assist in delivering forwards-compatibility
        with the :class:`_orm.WriteOnlyCollection` collection class.

        .. versionadded:: 2.0

        """
        self._add_all_impl(iterator)

    def add(self, item: _T) -> None:
        """Add an item to this :class:`_orm.AppenderQuery`.

        The given item will be persisted to the database in terms of
        the parent instance's collection on the next flush.

        This method is provided to assist in delivering forwards-compatibility
        with the :class:`_orm.WriteOnlyCollection` collection class.

        .. versionadded:: 2.0

        """
        self._add_all_impl([item])

    def extend(self, iterator: Iterable[_T]) -> None:
        """Add an iterable of items to this :class:`_orm.AppenderQuery`.

        The given items will be persisted to the database in terms of
        the parent instance's collection on the next flush.

        """
        self._add_all_impl(iterator)

    def append(self, item: _T) -> None:
        """Append an item to this :class:`_orm.AppenderQuery`.

        The given item will be persisted to the database in terms of
        the parent instance's collection on the next flush.

        """
        self._add_all_impl([item])

    def remove(self, item: _T) -> None:
        """Remove an item from this :class:`_orm.AppenderQuery`.

        The given item will be removed from the parent instance's collection on
        the next flush.

        """
        self._remove_impl(item)


class AppenderQuery(AppenderMixin[_T], Query[_T]):  # type: ignore[misc]
    """A dynamic query that supports basic collection storage operations.

    Methods on :class:`.AppenderQuery` include all methods of
    :class:`_orm.Query`, plus additional methods used for collection
    persistence.


    """


def mixin_user_query(cls: Any) -> type[AppenderMixin[Any]]:
    """Return a new class with AppenderQuery functionality layered over."""
    name = "Appender" + cls.__name__
    return type(name, (AppenderMixin, cls), {"query_class": cls})

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\sparsetools.py ===
from sympy.core.containers import Dict
from sympy.core.symbol import Dummy
from sympy.utilities.iterables import is_sequence
from sympy.utilities.misc import as_int, filldedent

from .sparse import MutableSparseMatrix as SparseMatrix


def _doktocsr(dok):
    """Converts a sparse matrix to Compressed Sparse Row (CSR) format.

    Parameters
    ==========

    A : contains non-zero elements sorted by key (row, column)
    JA : JA[i] is the column corresponding to A[i]
    IA : IA[i] contains the index in A for the first non-zero element
        of row[i]. Thus IA[i+1] - IA[i] gives number of non-zero
        elements row[i]. The length of IA is always 1 more than the
        number of rows in the matrix.

    Examples
    ========

    >>> from sympy.matrices.sparsetools import _doktocsr
    >>> from sympy import SparseMatrix, diag
    >>> m = SparseMatrix(diag(1, 2, 3))
    >>> m[2, 0] = -1
    >>> _doktocsr(m)
    [[1, 2, -1, 3], [0, 1, 0, 2], [0, 1, 2, 4], [3, 3]]

    """
    row, JA, A = [list(i) for i in zip(*dok.row_list())]
    IA = [0]*((row[0] if row else 0) + 1)
    for i, r in enumerate(row):
        IA.extend([i]*(r - row[i - 1]))  # if i = 0 nothing is extended
    IA.extend([len(A)]*(dok.rows - len(IA) + 1))
    shape = [dok.rows, dok.cols]
    return [A, JA, IA, shape]


def _csrtodok(csr):
    """Converts a CSR representation to DOK representation.

    Examples
    ========

    >>> from sympy.matrices.sparsetools import _csrtodok
    >>> _csrtodok([[5, 8, 3, 6], [0, 1, 2, 1], [0, 0, 2, 3, 4], [4, 3]])
    Matrix([
    [0, 0, 0],
    [5, 8, 0],
    [0, 0, 3],
    [0, 6, 0]])

    """
    smat = {}
    A, JA, IA, shape = csr
    for i in range(len(IA) - 1):
        indices = slice(IA[i], IA[i + 1])
        for l, m in zip(A[indices], JA[indices]):
            smat[i, m] = l
    return SparseMatrix(*shape, smat)


def banded(*args, **kwargs):
    """Returns a SparseMatrix from the given dictionary describing
    the diagonals of the matrix. The keys are positive for upper
    diagonals and negative for those below the main diagonal. The
    values may be:

    * expressions or single-argument functions,

    * lists or tuples of values,

    * matrices

    Unless dimensions are given, the size of the returned matrix will
    be large enough to contain the largest non-zero value provided.

    kwargs
    ======

    rows : rows of the resulting matrix; computed if
           not given.

    cols : columns of the resulting matrix; computed if
           not given.

    Examples
    ========

    >>> from sympy import banded, ones, Matrix
    >>> from sympy.abc import x

    If explicit values are given in tuples,
    the matrix will autosize to contain all values, otherwise
    a single value is filled onto the entire diagonal:

    >>> banded({1: (1, 2, 3), -1: (4, 5, 6), 0: x})
    Matrix([
    [x, 1, 0, 0],
    [4, x, 2, 0],
    [0, 5, x, 3],
    [0, 0, 6, x]])

    A function accepting a single argument can be used to fill the
    diagonal as a function of diagonal index (which starts at 0).
    The size (or shape) of the matrix must be given to obtain more
    than a 1x1 matrix:

    >>> s = lambda d: (1 + d)**2
    >>> banded(5, {0: s, 2: s, -2: 2})
    Matrix([
    [1, 0, 1,  0,  0],
    [0, 4, 0,  4,  0],
    [2, 0, 9,  0,  9],
    [0, 2, 0, 16,  0],
    [0, 0, 2,  0, 25]])

    The diagonal of matrices placed on a diagonal will coincide
    with the indicated diagonal:

    >>> vert = Matrix([1, 2, 3])
    >>> banded({0: vert}, cols=3)
    Matrix([
    [1, 0, 0],
    [2, 1, 0],
    [3, 2, 1],
    [0, 3, 2],
    [0, 0, 3]])

    >>> banded(4, {0: ones(2)})
    Matrix([
    [1, 1, 0, 0],
    [1, 1, 0, 0],
    [0, 0, 1, 1],
    [0, 0, 1, 1]])

    Errors are raised if the designated size will not hold
    all values an integral number of times. Here, the rows
    are designated as odd (but an even number is required to
    hold the off-diagonal 2x2 ones):

    >>> banded({0: 2, 1: ones(2)}, rows=5)
    Traceback (most recent call last):
    ...
    ValueError:
    sequence does not fit an integral number of times in the matrix

    And here, an even number of rows is given...but the square
    matrix has an even number of columns, too. As we saw
    in the previous example, an odd number is required:

    >>> banded(4, {0: 2, 1: ones(2)})  # trying to make 4x4 and cols must be odd
    Traceback (most recent call last):
    ...
    ValueError:
    sequence does not fit an integral number of times in the matrix

    A way around having to count rows is to enclosing matrix elements
    in a tuple and indicate the desired number of them to the right:

    >>> banded({0: 2, 2: (ones(2),)*3})
    Matrix([
    [2, 0, 1, 1, 0, 0, 0, 0],
    [0, 2, 1, 1, 0, 0, 0, 0],
    [0, 0, 2, 0, 1, 1, 0, 0],
    [0, 0, 0, 2, 1, 1, 0, 0],
    [0, 0, 0, 0, 2, 0, 1, 1],
    [0, 0, 0, 0, 0, 2, 1, 1]])

    An error will be raised if more than one value
    is written to a given entry. Here, the ones overlap
    with the main diagonal if they are placed on the
    first diagonal:

    >>> banded({0: (2,)*5, 1: (ones(2),)*3})
    Traceback (most recent call last):
    ...
    ValueError: collision at (1, 1)

    By placing a 0 at the bottom left of the 2x2 matrix of
    ones, the collision is avoided:

    >>> u2 = Matrix([
    ... [1, 1],
    ... [0, 1]])
    >>> banded({0: [2]*5, 1: [u2]*3})
    Matrix([
    [2, 1, 1, 0, 0, 0, 0],
    [0, 2, 1, 0, 0, 0, 0],
    [0, 0, 2, 1, 1, 0, 0],
    [0, 0, 0, 2, 1, 0, 0],
    [0, 0, 0, 0, 2, 1, 1],
    [0, 0, 0, 0, 0, 0, 1]])
    """
    try:
        if len(args) not in (1, 2, 3):
            raise TypeError
        if not isinstance(args[-1], (dict, Dict)):
            raise TypeError
        if len(args) == 1:
            rows = kwargs.get('rows', None)
            cols = kwargs.get('cols', None)
            if rows is not None:
                rows = as_int(rows)
            if cols is not None:
                cols = as_int(cols)
        elif len(args) == 2:
            rows = cols = as_int(args[0])
        else:
            rows, cols = map(as_int, args[:2])
        # fails with ValueError if any keys are not ints
        _ = all(as_int(k) for k in args[-1])
    except (ValueError, TypeError):
        raise TypeError(filldedent(
            '''unrecognized input to banded:
            expecting [[row,] col,] {int: value}'''))
    def rc(d):
        # return row,col coord of diagonal start
        r = -d if d < 0 else 0
        c = 0 if r else d
        return r, c
    smat = {}
    undone = []
    tba = Dummy()
    # first handle objects with size
    for d, v in args[-1].items():
        r, c = rc(d)
        # note: only list and tuple are recognized since this
        # will allow other Basic objects like Tuple
        # into the matrix if so desired
        if isinstance(v, (list, tuple)):
            extra = 0
            for i, vi in enumerate(v):
                i += extra
                if is_sequence(vi):
                    vi = SparseMatrix(vi)
                    smat[r + i, c + i] = vi
                    extra += min(vi.shape) - 1
                else:
                    smat[r + i, c + i] = vi
        elif is_sequence(v):
            v = SparseMatrix(v)
            rv, cv = v.shape
            if rows and cols:
                nr, xr = divmod(rows - r, rv)
                nc, xc = divmod(cols - c, cv)
                x = xr or xc
                do = min(nr, nc)
            elif rows:
                do, x = divmod(rows - r, rv)
            elif cols:
                do, x = divmod(cols - c, cv)
            else:
                do = 1
                x = 0
            if x:
                raise ValueError(filldedent('''
                    sequence does not fit an integral number of times
                    in the matrix'''))
            j = min(v.shape)
            for i in range(do):
                smat[r, c] = v
                r += j
                c += j
        elif v:
            smat[r, c] = tba
            undone.append((d, v))
    s = SparseMatrix(None, smat)  # to expand matrices
    smat = s.todok()
    # check for dim errors here
    if rows is not None and rows < s.rows:
        raise ValueError('Designated rows %s < needed %s' % (rows, s.rows))
    if cols is not None and cols < s.cols:
        raise ValueError('Designated cols %s < needed %s' % (cols, s.cols))
    if rows is cols is None:
        rows = s.rows
        cols = s.cols
    elif rows is not None and cols is None:
        cols = max(rows, s.cols)
    elif cols is not None and rows is None:
        rows = max(cols, s.rows)
    def update(i, j, v):
        # update smat and make sure there are
        # no collisions
        if v:
            if (i, j) in smat and smat[i, j] not in (tba, v):
                raise ValueError('collision at %s' % ((i, j),))
            smat[i, j] = v
    if undone:
        for d, vi in undone:
            r, c = rc(d)
            v = vi if callable(vi) else lambda _: vi
            i = 0
            while r + i < rows and c + i < cols:
                update(r + i, c + i, v(i))
                i += 1
    return SparseMatrix(rows, cols, smat)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\platformdirs\api.py ===
"""Base API."""

from __future__ import annotations

import os
from abc import ABC, abstractmethod
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from collections.abc import Iterator
    from typing import Literal


class PlatformDirsABC(ABC):  # noqa: PLR0904
    """Abstract base class for platform directories."""

    def __init__(  # noqa: PLR0913, PLR0917
        self,
        appname: str | None = None,
        appauthor: str | Literal[False] | None = None,
        version: str | None = None,
        roaming: bool = False,  # noqa: FBT001, FBT002
        multipath: bool = False,  # noqa: FBT001, FBT002
        opinion: bool = True,  # noqa: FBT001, FBT002
        ensure_exists: bool = False,  # noqa: FBT001, FBT002
    ) -> None:
        """
        Create a new platform directory.

        :param appname: See `appname`.
        :param appauthor: See `appauthor`.
        :param version: See `version`.
        :param roaming: See `roaming`.
        :param multipath: See `multipath`.
        :param opinion: See `opinion`.
        :param ensure_exists: See `ensure_exists`.

        """
        self.appname = appname  #: The name of application.
        self.appauthor = appauthor
        """
        The name of the app author or distributing body for this application.

        Typically, it is the owning company name. Defaults to `appname`. You may pass ``False`` to disable it.

        """
        self.version = version
        """
        An optional version path element to append to the path.

        You might want to use this if you want multiple versions of your app to be able to run independently. If used,
        this would typically be ``<major>.<minor>``.

        """
        self.roaming = roaming
        """
        Whether to use the roaming appdata directory on Windows.

        That means that for users on a Windows network setup for roaming profiles, this user data will be synced on
        login (see
        `here <https://technet.microsoft.com/en-us/library/cc766489(WS.10).aspx>`_).

        """
        self.multipath = multipath
        """
        An optional parameter which indicates that the entire list of data dirs should be returned.

        By default, the first item would only be returned.

        """
        self.opinion = opinion  #: A flag to indicating to use opinionated values.
        self.ensure_exists = ensure_exists
        """
        Optionally create the directory (and any missing parents) upon access if it does not exist.

        By default, no directories are created.

        """

    def _append_app_name_and_version(self, *base: str) -> str:
        params = list(base[1:])
        if self.appname:
            params.append(self.appname)
            if self.version:
                params.append(self.version)
        path = os.path.join(base[0], *params)  # noqa: PTH118
        self._optionally_create_directory(path)
        return path

    def _optionally_create_directory(self, path: str) -> None:
        if self.ensure_exists:
            Path(path).mkdir(parents=True, exist_ok=True)

    def _first_item_as_path_if_multipath(self, directory: str) -> Path:
        if self.multipath:
            # If multipath is True, the first path is returned.
            directory = directory.split(os.pathsep)[0]
        return Path(directory)

    @property
    @abstractmethod
    def user_data_dir(self) -> str:
        """:return: data directory tied to the user"""

    @property
    @abstractmethod
    def site_data_dir(self) -> str:
        """:return: data directory shared by users"""

    @property
    @abstractmethod
    def user_config_dir(self) -> str:
        """:return: config directory tied to the user"""

    @property
    @abstractmethod
    def site_config_dir(self) -> str:
        """:return: config directory shared by the users"""

    @property
    @abstractmethod
    def user_cache_dir(self) -> str:
        """:return: cache directory tied to the user"""

    @property
    @abstractmethod
    def site_cache_dir(self) -> str:
        """:return: cache directory shared by users"""

    @property
    @abstractmethod
    def user_state_dir(self) -> str:
        """:return: state directory tied to the user"""

    @property
    @abstractmethod
    def user_log_dir(self) -> str:
        """:return: log directory tied to the user"""

    @property
    @abstractmethod
    def user_documents_dir(self) -> str:
        """:return: documents directory tied to the user"""

    @property
    @abstractmethod
    def user_downloads_dir(self) -> str:
        """:return: downloads directory tied to the user"""

    @property
    @abstractmethod
    def user_pictures_dir(self) -> str:
        """:return: pictures directory tied to the user"""

    @property
    @abstractmethod
    def user_videos_dir(self) -> str:
        """:return: videos directory tied to the user"""

    @property
    @abstractmethod
    def user_music_dir(self) -> str:
        """:return: music directory tied to the user"""

    @property
    @abstractmethod
    def user_desktop_dir(self) -> str:
        """:return: desktop directory tied to the user"""

    @property
    @abstractmethod
    def user_runtime_dir(self) -> str:
        """:return: runtime directory tied to the user"""

    @property
    @abstractmethod
    def site_runtime_dir(self) -> str:
        """:return: runtime directory shared by users"""

    @property
    def user_data_path(self) -> Path:
        """:return: data path tied to the user"""
        return Path(self.user_data_dir)

    @property
    def site_data_path(self) -> Path:
        """:return: data path shared by users"""
        return Path(self.site_data_dir)

    @property
    def user_config_path(self) -> Path:
        """:return: config path tied to the user"""
        return Path(self.user_config_dir)

    @property
    def site_config_path(self) -> Path:
        """:return: config path shared by the users"""
        return Path(self.site_config_dir)

    @property
    def user_cache_path(self) -> Path:
        """:return: cache path tied to the user"""
        return Path(self.user_cache_dir)

    @property
    def site_cache_path(self) -> Path:
        """:return: cache path shared by users"""
        return Path(self.site_cache_dir)

    @property
    def user_state_path(self) -> Path:
        """:return: state path tied to the user"""
        return Path(self.user_state_dir)

    @property
    def user_log_path(self) -> Path:
        """:return: log path tied to the user"""
        return Path(self.user_log_dir)

    @property
    def user_documents_path(self) -> Path:
        """:return: documents a path tied to the user"""
        return Path(self.user_documents_dir)

    @property
    def user_downloads_path(self) -> Path:
        """:return: downloads path tied to the user"""
        return Path(self.user_downloads_dir)

    @property
    def user_pictures_path(self) -> Path:
        """:return: pictures path tied to the user"""
        return Path(self.user_pictures_dir)

    @property
    def user_videos_path(self) -> Path:
        """:return: videos path tied to the user"""
        return Path(self.user_videos_dir)

    @property
    def user_music_path(self) -> Path:
        """:return: music path tied to the user"""
        return Path(self.user_music_dir)

    @property
    def user_desktop_path(self) -> Path:
        """:return: desktop path tied to the user"""
        return Path(self.user_desktop_dir)

    @property
    def user_runtime_path(self) -> Path:
        """:return: runtime path tied to the user"""
        return Path(self.user_runtime_dir)

    @property
    def site_runtime_path(self) -> Path:
        """:return: runtime path shared by users"""
        return Path(self.site_runtime_dir)

    def iter_config_dirs(self) -> Iterator[str]:
        """:yield: all user and site configuration directories."""
        yield self.user_config_dir
        yield self.site_config_dir

    def iter_data_dirs(self) -> Iterator[str]:
        """:yield: all user and site data directories."""
        yield self.user_data_dir
        yield self.site_data_dir

    def iter_cache_dirs(self) -> Iterator[str]:
        """:yield: all user and site cache directories."""
        yield self.user_cache_dir
        yield self.site_cache_dir

    def iter_runtime_dirs(self) -> Iterator[str]:
        """:yield: all user and site runtime directories."""
        yield self.user_runtime_dir
        yield self.site_runtime_dir

    def iter_config_paths(self) -> Iterator[Path]:
        """:yield: all user and site configuration paths."""
        for path in self.iter_config_dirs():
            yield Path(path)

    def iter_data_paths(self) -> Iterator[Path]:
        """:yield: all user and site data paths."""
        for path in self.iter_data_dirs():
            yield Path(path)

    def iter_cache_paths(self) -> Iterator[Path]:
        """:yield: all user and site cache paths."""
        for path in self.iter_cache_dirs():
            yield Path(path)

    def iter_runtime_paths(self) -> Iterator[Path]:
        """:yield: all user and site runtime paths."""
        for path in self.iter_runtime_dirs():
            yield Path(path)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\expressions\special.py ===
from sympy.assumptions.ask import ask, Q
from sympy.core.relational import Eq
from sympy.core.singleton import S
from sympy.core.sympify import _sympify
from sympy.functions.special.tensor_functions import KroneckerDelta
from sympy.matrices.exceptions import NonInvertibleMatrixError
from .matexpr import MatrixExpr


class ZeroMatrix(MatrixExpr):
    """The Matrix Zero 0 - additive identity

    Examples
    ========

    >>> from sympy import MatrixSymbol, ZeroMatrix
    >>> A = MatrixSymbol('A', 3, 5)
    >>> Z = ZeroMatrix(3, 5)
    >>> A + Z
    A
    >>> Z*A.T
    0
    """
    is_ZeroMatrix = True

    def __new__(cls, m, n):
        m, n = _sympify(m), _sympify(n)
        cls._check_dim(m)
        cls._check_dim(n)

        return super().__new__(cls, m, n)

    @property
    def shape(self):
        return (self.args[0], self.args[1])

    def _eval_power(self, exp):
        # exp = -1, 0, 1 are already handled at this stage
        if (exp < 0) == True:
            raise NonInvertibleMatrixError("Matrix det == 0; not invertible")
        return self

    def _eval_transpose(self):
        return ZeroMatrix(self.cols, self.rows)

    def _eval_adjoint(self):
        return ZeroMatrix(self.cols, self.rows)

    def _eval_trace(self):
        return S.Zero

    def _eval_determinant(self):
        return S.Zero

    def _eval_inverse(self):
        raise NonInvertibleMatrixError("Matrix det == 0; not invertible.")

    def _eval_as_real_imag(self):
        return (self, self)

    def _eval_conjugate(self):
        return self

    def _entry(self, i, j, **kwargs):
        return S.Zero


class GenericZeroMatrix(ZeroMatrix):
    """
    A zero matrix without a specified shape

    This exists primarily so MatAdd() with no arguments can return something
    meaningful.
    """
    def __new__(cls):
        # super(ZeroMatrix, cls) instead of super(GenericZeroMatrix, cls)
        # because ZeroMatrix.__new__ doesn't have the same signature
        return super(ZeroMatrix, cls).__new__(cls)

    @property
    def rows(self):
        raise TypeError("GenericZeroMatrix does not have a specified shape")

    @property
    def cols(self):
        raise TypeError("GenericZeroMatrix does not have a specified shape")

    @property
    def shape(self):
        raise TypeError("GenericZeroMatrix does not have a specified shape")

    # Avoid Matrix.__eq__ which might call .shape
    def __eq__(self, other):
        return isinstance(other, GenericZeroMatrix)

    def __ne__(self, other):
        return not (self == other)

    def __hash__(self):
        return super().__hash__()



class Identity(MatrixExpr):
    """The Matrix Identity I - multiplicative identity

    Examples
    ========

    >>> from sympy import Identity, MatrixSymbol
    >>> A = MatrixSymbol('A', 3, 5)
    >>> I = Identity(3)
    >>> I*A
    A
    """

    is_Identity = True

    def __new__(cls, n):
        n = _sympify(n)
        cls._check_dim(n)

        return super().__new__(cls, n)

    @property
    def rows(self):
        return self.args[0]

    @property
    def cols(self):
        return self.args[0]

    @property
    def shape(self):
        return (self.args[0], self.args[0])

    @property
    def is_square(self):
        return True

    def _eval_transpose(self):
        return self

    def _eval_trace(self):
        return self.rows

    def _eval_inverse(self):
        return self

    def _eval_as_real_imag(self):
        return (self, ZeroMatrix(*self.shape))

    def _eval_conjugate(self):
        return self

    def _eval_adjoint(self):
        return self

    def _entry(self, i, j, **kwargs):
        eq = Eq(i, j)
        if eq is S.true:
            return S.One
        elif eq is S.false:
            return S.Zero
        return KroneckerDelta(i, j, (0, self.cols-1))

    def _eval_determinant(self):
        return S.One

    def _eval_power(self, exp):
        return self


class GenericIdentity(Identity):
    """
    An identity matrix without a specified shape

    This exists primarily so MatMul() with no arguments can return something
    meaningful.
    """
    def __new__(cls):
        # super(Identity, cls) instead of super(GenericIdentity, cls) because
        # Identity.__new__ doesn't have the same signature
        return super(Identity, cls).__new__(cls)

    @property
    def rows(self):
        raise TypeError("GenericIdentity does not have a specified shape")

    @property
    def cols(self):
        raise TypeError("GenericIdentity does not have a specified shape")

    @property
    def shape(self):
        raise TypeError("GenericIdentity does not have a specified shape")

    @property
    def is_square(self):
        return True

    # Avoid Matrix.__eq__ which might call .shape
    def __eq__(self, other):
        return isinstance(other, GenericIdentity)

    def __ne__(self, other):
        return not (self == other)

    def __hash__(self):
        return super().__hash__()


class OneMatrix(MatrixExpr):
    """
    Matrix whose all entries are ones.
    """
    def __new__(cls, m, n, evaluate=False):
        m, n = _sympify(m), _sympify(n)
        cls._check_dim(m)
        cls._check_dim(n)

        if evaluate:
            condition = Eq(m, 1) & Eq(n, 1)
            if condition == True:
                return Identity(1)

        obj = super().__new__(cls, m, n)
        return obj

    @property
    def shape(self):
        return self._args

    @property
    def is_Identity(self):
        return self._is_1x1() == True

    def as_explicit(self):
        from sympy.matrices.immutable import ImmutableDenseMatrix
        return ImmutableDenseMatrix.ones(*self.shape)

    def doit(self, **hints):
        args = self.args
        if hints.get('deep', True):
            args = [a.doit(**hints) for a in args]
        return self.func(*args, evaluate=True)

    def _eval_power(self, exp):
        # exp = -1, 0, 1 are already handled at this stage
        if self._is_1x1() == True:
            return Identity(1)
        if (exp < 0) == True:
            raise NonInvertibleMatrixError("Matrix det == 0; not invertible")
        if ask(Q.integer(exp)):
            return self.shape[0] ** (exp - 1) * OneMatrix(*self.shape)
        return super()._eval_power(exp)

    def _eval_transpose(self):
        return OneMatrix(self.cols, self.rows)

    def _eval_adjoint(self):
        return OneMatrix(self.cols, self.rows)

    def _eval_trace(self):
        return S.One*self.rows

    def _is_1x1(self):
        """Returns true if the matrix is known to be 1x1"""
        shape = self.shape
        return Eq(shape[0], 1) & Eq(shape[1], 1)

    def _eval_determinant(self):
        condition = self._is_1x1()
        if condition == True:
            return S.One
        elif condition == False:
            return S.Zero
        else:
            from sympy.matrices.expressions.determinant import Determinant
            return Determinant(self)

    def _eval_inverse(self):
        condition = self._is_1x1()
        if condition == True:
            return Identity(1)
        elif condition == False:
            raise NonInvertibleMatrixError("Matrix det == 0; not invertible.")
        else:
            from .inverse import Inverse
            return Inverse(self)

    def _eval_as_real_imag(self):
        return (self, ZeroMatrix(*self.shape))

    def _eval_conjugate(self):
        return self

    def _entry(self, i, j, **kwargs):
        return S.One

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_core\_tests\test_windows.py ===
from __future__ import annotations

import os
import sys
import tempfile
from contextlib import contextmanager
from typing import TYPE_CHECKING
from unittest.mock import create_autospec

import pytest

on_windows = os.name == "nt"
# Mark all the tests in this file as being windows-only
pytestmark = pytest.mark.skipif(not on_windows, reason="windows only")

assert (
    sys.platform == "win32" or not TYPE_CHECKING
)  # Skip type checking when not on Windows

from ... import _core, sleep
from ...testing import wait_all_tasks_blocked
from .tutil import gc_collect_harder, restore_unraisablehook, slow

if TYPE_CHECKING:
    from collections.abc import Generator
    from io import BufferedWriter

if on_windows:
    from .._windows_cffi import (
        INVALID_HANDLE_VALUE,
        FileFlags,
        Handle,
        ffi,
        kernel32,
        raise_winerror,
    )


def test_winerror(monkeypatch: pytest.MonkeyPatch) -> None:
    mock = create_autospec(ffi.getwinerror)
    monkeypatch.setattr(ffi, "getwinerror", mock)

    # Returning none = no error, should not happen.
    mock.return_value = None
    with pytest.raises(RuntimeError, match=r"^No error set\?$"):
        raise_winerror()
    mock.assert_called_once_with()
    mock.reset_mock()

    with pytest.raises(RuntimeError, match=r"^No error set\?$"):
        raise_winerror(38)
    mock.assert_called_once_with(38)
    mock.reset_mock()

    mock.return_value = (12, "test error")
    with pytest.raises(
        OSError,
        match=r"^\[WinError 12\] test error: 'file_1' -> 'file_2'$",
    ) as exc:
        raise_winerror(filename="file_1", filename2="file_2")
    mock.assert_called_once_with()
    mock.reset_mock()
    assert exc.value.winerror == 12
    assert exc.value.strerror == "test error"
    assert exc.value.filename == "file_1"
    assert exc.value.filename2 == "file_2"

    # With an explicit number passed in, it overrides what getwinerror() returns.
    with pytest.raises(
        OSError,
        match=r"^\[WinError 18\] test error: 'a/file' -> 'b/file'$",
    ) as exc:
        raise_winerror(18, filename="a/file", filename2="b/file")
    mock.assert_called_once_with(18)
    mock.reset_mock()
    assert exc.value.winerror == 18
    assert exc.value.strerror == "test error"
    assert exc.value.filename == "a/file"
    assert exc.value.filename2 == "b/file"


# The undocumented API that this is testing should be changed to stop using
# UnboundedQueue (or just removed until we have time to redo it), but until
# then we filter out the warning.
@pytest.mark.filterwarnings("ignore:.*UnboundedQueue:trio.TrioDeprecationWarning")
async def test_completion_key_listen() -> None:
    from .. import _io_windows

    async def post(key: int) -> None:
        iocp = Handle(ffi.cast("HANDLE", _core.current_iocp()))
        for i in range(10):
            print("post", i)
            if i % 3 == 0:
                await _core.checkpoint()
            success = kernel32.PostQueuedCompletionStatus(iocp, i, key, ffi.NULL)
            assert success

    with _core.monitor_completion_key() as (key, queue):
        async with _core.open_nursery() as nursery:
            nursery.start_soon(post, key)
            i = 0
            print("loop")
            async for batch in queue:  # pragma: no branch
                print("got some", batch)
                for info in batch:
                    assert isinstance(info, _io_windows.CompletionKeyEventInfo)
                    assert info.lpOverlapped == 0
                    assert info.dwNumberOfBytesTransferred == i
                    i += 1
                if i == 10:
                    break
            print("end loop")


async def test_readinto_overlapped() -> None:
    data = b"1" * 1024 + b"2" * 1024 + b"3" * 1024 + b"4" * 1024
    buffer = bytearray(len(data))

    with tempfile.TemporaryDirectory() as tdir:
        tfile = os.path.join(tdir, "numbers.txt")
        with open(  # noqa: ASYNC230  # This is a test, synchronous is ok
            tfile,
            "wb",
        ) as fp:
            fp.write(data)
            fp.flush()

        rawname = tfile.encode("utf-16le") + b"\0\0"
        rawname_buf = ffi.from_buffer(rawname)
        handle = kernel32.CreateFileW(
            ffi.cast("LPCWSTR", rawname_buf),
            FileFlags.GENERIC_READ,
            FileFlags.FILE_SHARE_READ,
            ffi.NULL,  # no security attributes
            FileFlags.OPEN_EXISTING,
            FileFlags.FILE_FLAG_OVERLAPPED,
            ffi.NULL,  # no template file
        )
        if handle == INVALID_HANDLE_VALUE:  # pragma: no cover
            raise_winerror()

        try:
            with memoryview(buffer) as buffer_view:

                async def read_region(start: int, end: int) -> None:
                    await _core.readinto_overlapped(
                        handle,
                        buffer_view[start:end],
                        start,
                    )

                _core.register_with_iocp(handle)
                async with _core.open_nursery() as nursery:
                    for start in range(0, 4096, 512):
                        nursery.start_soon(read_region, start, start + 512)

                assert buffer == data

                with pytest.raises((BufferError, TypeError)):
                    await _core.readinto_overlapped(handle, b"immutable")
        finally:
            kernel32.CloseHandle(handle)


@contextmanager
def pipe_with_overlapped_read() -> Generator[tuple[BufferedWriter, int], None, None]:
    import msvcrt
    from asyncio.windows_utils import pipe

    read_handle, write_handle = pipe(overlapped=(True, False))
    try:
        write_fd = msvcrt.open_osfhandle(write_handle, 0)
        yield os.fdopen(write_fd, "wb", closefd=False), read_handle
    finally:
        kernel32.CloseHandle(Handle(ffi.cast("HANDLE", read_handle)))
        kernel32.CloseHandle(Handle(ffi.cast("HANDLE", write_handle)))


@restore_unraisablehook()
def test_forgot_to_register_with_iocp() -> None:
    with pipe_with_overlapped_read() as (write_fp, read_handle):
        with write_fp:
            write_fp.write(b"test\n")

        left_run_yet = False

        async def main() -> None:
            target = bytearray(1)
            try:
                async with _core.open_nursery() as nursery:
                    nursery.start_soon(
                        _core.readinto_overlapped,
                        read_handle,
                        target,
                        name="xyz",
                    )
                    await wait_all_tasks_blocked()
                    nursery.cancel_scope.cancel()
            finally:
                # Run loop is exited without unwinding running tasks, so
                # we don't get here until the main() coroutine is GC'ed
                assert left_run_yet

        with pytest.raises(_core.TrioInternalError) as exc_info:
            _core.run(main)
        left_run_yet = True
        assert "Failed to cancel overlapped I/O in xyz " in str(exc_info.value)
        assert "forget to call register_with_iocp()?" in str(exc_info.value)

        # Make sure the Nursery.__del__ assertion about dangling children
        # gets put with the correct test
        del exc_info
        gc_collect_harder()


@slow
async def test_too_late_to_cancel() -> None:
    import time

    with pipe_with_overlapped_read() as (write_fp, read_handle):
        _core.register_with_iocp(read_handle)
        target = bytearray(6)
        async with _core.open_nursery() as nursery:
            # Start an async read in the background
            nursery.start_soon(_core.readinto_overlapped, read_handle, target)
            await wait_all_tasks_blocked()

            # Synchronous write to the other end of the pipe
            with write_fp:
                write_fp.write(b"test1\ntest2\n")

            # Note: not trio.sleep! We're making sure the OS level
            # ReadFile completes, before Trio has a chance to execute
            # another checkpoint and notice it completed.
            time.sleep(1)  # noqa: ASYNC251
            nursery.cancel_scope.cancel()
        assert target[:6] == b"test1\n"

        # Do another I/O to make sure we've actually processed the
        # fallback completion that was posted when CancelIoEx failed.
        assert await _core.readinto_overlapped(read_handle, target) == 6
        assert target[:6] == b"test2\n"


def test_lsp_that_hooks_select_gives_good_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from .. import _io_windows
    from .._windows_cffi import CData, WSAIoctls, _handle

    def patched_get_underlying(
        sock: int | CData,
        *,
        which: int = WSAIoctls.SIO_BASE_HANDLE,
    ) -> CData:
        if hasattr(sock, "fileno"):  # pragma: no branch
            sock = sock.fileno()
        if which == WSAIoctls.SIO_BSP_HANDLE_SELECT:
            return _handle(sock + 1)
        else:
            return _handle(sock)

    monkeypatch.setattr(_io_windows, "_get_underlying_socket", patched_get_underlying)
    with pytest.raises(
        RuntimeError,
        match="SIO_BASE_HANDLE and SIO_BSP_HANDLE_SELECT differ",
    ):
        _core.run(sleep, 0)


def test_lsp_that_completely_hides_base_socket_gives_good_error(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # This tests behavior with an LSP that fails SIO_BASE_HANDLE and returns
    # self for SIO_BSP_HANDLE_SELECT (like Komodia), but also returns
    # self for SIO_BSP_HANDLE_POLL. No known LSP does this, but we want to
    # make sure we get an error rather than an infinite loop.

    from .. import _io_windows
    from .._windows_cffi import CData, WSAIoctls, _handle

    def patched_get_underlying(
        sock: int | CData,
        *,
        which: int = WSAIoctls.SIO_BASE_HANDLE,
    ) -> CData:
        if hasattr(sock, "fileno"):  # pragma: no branch
            sock = sock.fileno()
        if which == WSAIoctls.SIO_BASE_HANDLE:
            raise OSError("nope")
        else:
            return _handle(sock)

    monkeypatch.setattr(_io_windows, "_get_underlying_socket", patched_get_underlying)
    with pytest.raises(
        RuntimeError,
        match="SIO_BASE_HANDLE failed and SIO_BSP_HANDLE_POLL didn't return a diff",
    ):
        _core.run(sleep, 0)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\flask\__init__.py ===
from __future__ import annotations

import typing as t

from . import json as json
from .app import Flask as Flask
from .blueprints import Blueprint as Blueprint
from .config import Config as Config
from .ctx import after_this_request as after_this_request
from .ctx import copy_current_request_context as copy_current_request_context
from .ctx import has_app_context as has_app_context
from .ctx import has_request_context as has_request_context
from .globals import current_app as current_app
from .globals import g as g
from .globals import request as request
from .globals import session as session
from .helpers import abort as abort
from .helpers import flash as flash
from .helpers import get_flashed_messages as get_flashed_messages
from .helpers import get_template_attribute as get_template_attribute
from .helpers import make_response as make_response
from .helpers import redirect as redirect
from .helpers import send_file as send_file
from .helpers import send_from_directory as send_from_directory
from .helpers import stream_with_context as stream_with_context
from .helpers import url_for as url_for
from .json import jsonify as jsonify
from .signals import appcontext_popped as appcontext_popped
from .signals import appcontext_pushed as appcontext_pushed
from .signals import appcontext_tearing_down as appcontext_tearing_down
from .signals import before_render_template as before_render_template
from .signals import got_request_exception as got_request_exception
from .signals import message_flashed as message_flashed
from .signals import request_finished as request_finished
from .signals import request_started as request_started
from .signals import request_tearing_down as request_tearing_down
from .signals import template_rendered as template_rendered
from .templating import render_template as render_template
from .templating import render_template_string as render_template_string
from .templating import stream_template as stream_template
from .templating import stream_template_string as stream_template_string
from .wrappers import Request as Request
from .wrappers import Response as Response

if not t.TYPE_CHECKING:

    def __getattr__(name: str) -> t.Any:
        if name == "__version__":
            import importlib.metadata
            import warnings

            warnings.warn(
                "The '__version__' attribute is deprecated and will be removed in"
                " Flask 3.2. Use feature detection or"
                " 'importlib.metadata.version(\"flask\")' instead.",
                DeprecationWarning,
                stacklevel=2,
            )
            return importlib.metadata.version("flask")

        raise AttributeError(name)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\flask\testing.py ===
from __future__ import annotations

import importlib.metadata
import typing as t
from contextlib import contextmanager
from contextlib import ExitStack
from copy import copy
from types import TracebackType
from urllib.parse import urlsplit

import werkzeug.test
from click.testing import CliRunner
from click.testing import Result
from werkzeug.test import Client
from werkzeug.wrappers import Request as BaseRequest

from .cli import ScriptInfo
from .sessions import SessionMixin

if t.TYPE_CHECKING:  # pragma: no cover
    from _typeshed.wsgi import WSGIEnvironment
    from werkzeug.test import TestResponse

    from .app import Flask


class EnvironBuilder(werkzeug.test.EnvironBuilder):
    """An :class:`~werkzeug.test.EnvironBuilder`, that takes defaults from the
    application.

    :param app: The Flask application to configure the environment from.
    :param path: URL path being requested.
    :param base_url: Base URL where the app is being served, which
        ``path`` is relative to. If not given, built from
        :data:`PREFERRED_URL_SCHEME`, ``subdomain``,
        :data:`SERVER_NAME`, and :data:`APPLICATION_ROOT`.
    :param subdomain: Subdomain name to append to :data:`SERVER_NAME`.
    :param url_scheme: Scheme to use instead of
        :data:`PREFERRED_URL_SCHEME`.
    :param json: If given, this is serialized as JSON and passed as
        ``data``. Also defaults ``content_type`` to
        ``application/json``.
    :param args: other positional arguments passed to
        :class:`~werkzeug.test.EnvironBuilder`.
    :param kwargs: other keyword arguments passed to
        :class:`~werkzeug.test.EnvironBuilder`.
    """

    def __init__(
        self,
        app: Flask,
        path: str = "/",
        base_url: str | None = None,
        subdomain: str | None = None,
        url_scheme: str | None = None,
        *args: t.Any,
        **kwargs: t.Any,
    ) -> None:
        assert not (base_url or subdomain or url_scheme) or (
            base_url is not None
        ) != bool(subdomain or url_scheme), (
            'Cannot pass "subdomain" or "url_scheme" with "base_url".'
        )

        if base_url is None:
            http_host = app.config.get("SERVER_NAME") or "localhost"
            app_root = app.config["APPLICATION_ROOT"]

            if subdomain:
                http_host = f"{subdomain}.{http_host}"

            if url_scheme is None:
                url_scheme = app.config["PREFERRED_URL_SCHEME"]

            url = urlsplit(path)
            base_url = (
                f"{url.scheme or url_scheme}://{url.netloc or http_host}"
                f"/{app_root.lstrip('/')}"
            )
            path = url.path

            if url.query:
                path = f"{path}?{url.query}"

        self.app = app
        super().__init__(path, base_url, *args, **kwargs)

    def json_dumps(self, obj: t.Any, **kwargs: t.Any) -> str:  # type: ignore
        """Serialize ``obj`` to a JSON-formatted string.

        The serialization will be configured according to the config associated
        with this EnvironBuilder's ``app``.
        """
        return self.app.json.dumps(obj, **kwargs)


_werkzeug_version = ""


def _get_werkzeug_version() -> str:
    global _werkzeug_version

    if not _werkzeug_version:
        _werkzeug_version = importlib.metadata.version("werkzeug")

    return _werkzeug_version


class FlaskClient(Client):
    """Works like a regular Werkzeug test client but has knowledge about
    Flask's contexts to defer the cleanup of the request context until
    the end of a ``with`` block. For general information about how to
    use this class refer to :class:`werkzeug.test.Client`.

    .. versionchanged:: 0.12
       `app.test_client()` includes preset default environment, which can be
       set after instantiation of the `app.test_client()` object in
       `client.environ_base`.

    Basic usage is outlined in the :doc:`/testing` chapter.
    """

    application: Flask

    def __init__(self, *args: t.Any, **kwargs: t.Any) -> None:
        super().__init__(*args, **kwargs)
        self.preserve_context = False
        self._new_contexts: list[t.ContextManager[t.Any]] = []
        self._context_stack = ExitStack()
        self.environ_base = {
            "REMOTE_ADDR": "127.0.0.1",
            "HTTP_USER_AGENT": f"Werkzeug/{_get_werkzeug_version()}",
        }

    @contextmanager
    def session_transaction(
        self, *args: t.Any, **kwargs: t.Any
    ) -> t.Iterator[SessionMixin]:
        """When used in combination with a ``with`` statement this opens a
        session transaction.  This can be used to modify the session that
        the test client uses.  Once the ``with`` block is left the session is
        stored back.

        ::

            with client.session_transaction() as session:
                session['value'] = 42

        Internally this is implemented by going through a temporary test
        request context and since session handling could depend on
        request variables this function accepts the same arguments as
        :meth:`~flask.Flask.test_request_context` which are directly
        passed through.
        """
        if self._cookies is None:
            raise TypeError(
                "Cookies are disabled. Create a client with 'use_cookies=True'."
            )

        app = self.application
        ctx = app.test_request_context(*args, **kwargs)
        self._add_cookies_to_wsgi(ctx.request.environ)

        with ctx:
            sess = app.session_interface.open_session(app, ctx.request)

        if sess is None:
            raise RuntimeError("Session backend did not open a session.")

        yield sess
        resp = app.response_class()

        if app.session_interface.is_null_session(sess):
            return

        with ctx:
            app.session_interface.save_session(app, sess, resp)

        self._update_cookies_from_response(
            ctx.request.host.partition(":")[0],
            ctx.request.path,
            resp.headers.getlist("Set-Cookie"),
        )

    def _copy_environ(self, other: WSGIEnvironment) -> WSGIEnvironment:
        out = {**self.environ_base, **other}

        if self.preserve_context:
            out["werkzeug.debug.preserve_context"] = self._new_contexts.append

        return out

    def _request_from_builder_args(
        self, args: tuple[t.Any, ...], kwargs: dict[str, t.Any]
    ) -> BaseRequest:
        kwargs["environ_base"] = self._copy_environ(kwargs.get("environ_base", {}))
        builder = EnvironBuilder(self.application, *args, **kwargs)

        try:
            return builder.get_request()
        finally:
            builder.close()

    def open(
        self,
        *args: t.Any,
        buffered: bool = False,
        follow_redirects: bool = False,
        **kwargs: t.Any,
    ) -> TestResponse:
        if args and isinstance(
            args[0], (werkzeug.test.EnvironBuilder, dict, BaseRequest)
        ):
            if isinstance(args[0], werkzeug.test.EnvironBuilder):
                builder = copy(args[0])
                builder.environ_base = self._copy_environ(builder.environ_base or {})  # type: ignore[arg-type]
                request = builder.get_request()
            elif isinstance(args[0], dict):
                request = EnvironBuilder.from_environ(
                    args[0], app=self.application, environ_base=self._copy_environ({})
                ).get_request()
            else:
                # isinstance(args[0], BaseRequest)
                request = copy(args[0])
                request.environ = self._copy_environ(request.environ)
        else:
            # request is None
            request = self._request_from_builder_args(args, kwargs)

        # Pop any previously preserved contexts. This prevents contexts
        # from being preserved across redirects or multiple requests
        # within a single block.
        self._context_stack.close()

        response = super().open(
            request,
            buffered=buffered,
            follow_redirects=follow_redirects,
        )
        response.json_module = self.application.json  # type: ignore[assignment]

        # Re-push contexts that were preserved during the request.
        while self._new_contexts:
            cm = self._new_contexts.pop()
            self._context_stack.enter_context(cm)

        return response

    def __enter__(self) -> FlaskClient:
        if self.preserve_context:
            raise RuntimeError("Cannot nest client invocations")
        self.preserve_context = True
        return self

    def __exit__(
        self,
        exc_type: type | None,
        exc_value: BaseException | None,
        tb: TracebackType | None,
    ) -> None:
        self.preserve_context = False
        self._context_stack.close()


class FlaskCliRunner(CliRunner):
    """A :class:`~click.testing.CliRunner` for testing a Flask app's
    CLI commands. Typically created using
    :meth:`~flask.Flask.test_cli_runner`. See :ref:`testing-cli`.
    """

    def __init__(self, app: Flask, **kwargs: t.Any) -> None:
        self.app = app
        super().__init__(**kwargs)

    def invoke(  # type: ignore
        self, cli: t.Any = None, args: t.Any = None, **kwargs: t.Any
    ) -> Result:
        """Invokes a CLI command in an isolated environment. See
        :meth:`CliRunner.invoke <click.testing.CliRunner.invoke>` for
        full method documentation. See :ref:`testing-cli` for examples.

        If the ``obj`` argument is not given, passes an instance of
        :class:`~flask.cli.ScriptInfo` that knows how to load the Flask
        app being tested.

        :param cli: Command object to invoke. Default is the app's
            :attr:`~flask.app.Flask.cli` group.
        :param args: List of strings to invoke the command with.

        :return: a :class:`~click.testing.Result` object.
        """
        if cli is None:
            cli = self.app.cli

        if "obj" not in kwargs:
            kwargs["obj"] = ScriptInfo(create_app=lambda: self.app)

        return super().invoke(cli, args, **kwargs)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\logging.py ===
import logging
from datetime import datetime
from logging import Handler, LogRecord
from pathlib import Path
from types import ModuleType
from typing import ClassVar, Iterable, List, Optional, Type, Union

from pip._vendor.rich._null_file import NullFile

from . import get_console
from ._log_render import FormatTimeCallable, LogRender
from .console import Console, ConsoleRenderable
from .highlighter import Highlighter, ReprHighlighter
from .text import Text
from .traceback import Traceback


class RichHandler(Handler):
    """A logging handler that renders output with Rich. The time / level / message and file are displayed in columns.
    The level is color coded, and the message is syntax highlighted.

    Note:
        Be careful when enabling console markup in log messages if you have configured logging for libraries not
        under your control. If a dependency writes messages containing square brackets, it may not produce the intended output.

    Args:
        level (Union[int, str], optional): Log level. Defaults to logging.NOTSET.
        console (:class:`~rich.console.Console`, optional): Optional console instance to write logs.
            Default will use a global console instance writing to stdout.
        show_time (bool, optional): Show a column for the time. Defaults to True.
        omit_repeated_times (bool, optional): Omit repetition of the same time. Defaults to True.
        show_level (bool, optional): Show a column for the level. Defaults to True.
        show_path (bool, optional): Show the path to the original log call. Defaults to True.
        enable_link_path (bool, optional): Enable terminal link of path column to file. Defaults to True.
        highlighter (Highlighter, optional): Highlighter to style log messages, or None to use ReprHighlighter. Defaults to None.
        markup (bool, optional): Enable console markup in log messages. Defaults to False.
        rich_tracebacks (bool, optional): Enable rich tracebacks with syntax highlighting and formatting. Defaults to False.
        tracebacks_width (Optional[int], optional): Number of characters used to render tracebacks, or None for full width. Defaults to None.
        tracebacks_code_width (int, optional): Number of code characters used to render tracebacks, or None for full width. Defaults to 88.
        tracebacks_extra_lines (int, optional): Additional lines of code to render tracebacks, or None for full width. Defaults to None.
        tracebacks_theme (str, optional): Override pygments theme used in traceback.
        tracebacks_word_wrap (bool, optional): Enable word wrapping of long tracebacks lines. Defaults to True.
        tracebacks_show_locals (bool, optional): Enable display of locals in tracebacks. Defaults to False.
        tracebacks_suppress (Sequence[Union[str, ModuleType]]): Optional sequence of modules or paths to exclude from traceback.
        tracebacks_max_frames (int, optional): Optional maximum number of frames returned by traceback.
        locals_max_length (int, optional): Maximum length of containers before abbreviating, or None for no abbreviation.
            Defaults to 10.
        locals_max_string (int, optional): Maximum length of string before truncating, or None to disable. Defaults to 80.
        log_time_format (Union[str, TimeFormatterCallable], optional): If ``log_time`` is enabled, either string for strftime or callable that formats the time. Defaults to "[%x %X] ".
        keywords (List[str], optional): List of words to highlight instead of ``RichHandler.KEYWORDS``.
    """

    KEYWORDS: ClassVar[Optional[List[str]]] = [
        "GET",
        "POST",
        "HEAD",
        "PUT",
        "DELETE",
        "OPTIONS",
        "TRACE",
        "PATCH",
    ]
    HIGHLIGHTER_CLASS: ClassVar[Type[Highlighter]] = ReprHighlighter

    def __init__(
        self,
        level: Union[int, str] = logging.NOTSET,
        console: Optional[Console] = None,
        *,
        show_time: bool = True,
        omit_repeated_times: bool = True,
        show_level: bool = True,
        show_path: bool = True,
        enable_link_path: bool = True,
        highlighter: Optional[Highlighter] = None,
        markup: bool = False,
        rich_tracebacks: bool = False,
        tracebacks_width: Optional[int] = None,
        tracebacks_code_width: int = 88,
        tracebacks_extra_lines: int = 3,
        tracebacks_theme: Optional[str] = None,
        tracebacks_word_wrap: bool = True,
        tracebacks_show_locals: bool = False,
        tracebacks_suppress: Iterable[Union[str, ModuleType]] = (),
        tracebacks_max_frames: int = 100,
        locals_max_length: int = 10,
        locals_max_string: int = 80,
        log_time_format: Union[str, FormatTimeCallable] = "[%x %X]",
        keywords: Optional[List[str]] = None,
    ) -> None:
        super().__init__(level=level)
        self.console = console or get_console()
        self.highlighter = highlighter or self.HIGHLIGHTER_CLASS()
        self._log_render = LogRender(
            show_time=show_time,
            show_level=show_level,
            show_path=show_path,
            time_format=log_time_format,
            omit_repeated_times=omit_repeated_times,
            level_width=None,
        )
        self.enable_link_path = enable_link_path
        self.markup = markup
        self.rich_tracebacks = rich_tracebacks
        self.tracebacks_width = tracebacks_width
        self.tracebacks_extra_lines = tracebacks_extra_lines
        self.tracebacks_theme = tracebacks_theme
        self.tracebacks_word_wrap = tracebacks_word_wrap
        self.tracebacks_show_locals = tracebacks_show_locals
        self.tracebacks_suppress = tracebacks_suppress
        self.tracebacks_max_frames = tracebacks_max_frames
        self.tracebacks_code_width = tracebacks_code_width
        self.locals_max_length = locals_max_length
        self.locals_max_string = locals_max_string
        self.keywords = keywords

    def get_level_text(self, record: LogRecord) -> Text:
        """Get the level name from the record.

        Args:
            record (LogRecord): LogRecord instance.

        Returns:
            Text: A tuple of the style and level name.
        """
        level_name = record.levelname
        level_text = Text.styled(
            level_name.ljust(8), f"logging.level.{level_name.lower()}"
        )
        return level_text

    def emit(self, record: LogRecord) -> None:
        """Invoked by logging."""
        message = self.format(record)
        traceback = None
        if (
            self.rich_tracebacks
            and record.exc_info
            and record.exc_info != (None, None, None)
        ):
            exc_type, exc_value, exc_traceback = record.exc_info
            assert exc_type is not None
            assert exc_value is not None
            traceback = Traceback.from_exception(
                exc_type,
                exc_value,
                exc_traceback,
                width=self.tracebacks_width,
                code_width=self.tracebacks_code_width,
                extra_lines=self.tracebacks_extra_lines,
                theme=self.tracebacks_theme,
                word_wrap=self.tracebacks_word_wrap,
                show_locals=self.tracebacks_show_locals,
                locals_max_length=self.locals_max_length,
                locals_max_string=self.locals_max_string,
                suppress=self.tracebacks_suppress,
                max_frames=self.tracebacks_max_frames,
            )
            message = record.getMessage()
            if self.formatter:
                record.message = record.getMessage()
                formatter = self.formatter
                if hasattr(formatter, "usesTime") and formatter.usesTime():
                    record.asctime = formatter.formatTime(record, formatter.datefmt)
                message = formatter.formatMessage(record)

        message_renderable = self.render_message(record, message)
        log_renderable = self.render(
            record=record, traceback=traceback, message_renderable=message_renderable
        )
        if isinstance(self.console.file, NullFile):
            # Handles pythonw, where stdout/stderr are null, and we return NullFile
            # instance from Console.file. In this case, we still want to make a log record
            # even though we won't be writing anything to a file.
            self.handleError(record)
        else:
            try:
                self.console.print(log_renderable)
            except Exception:
                self.handleError(record)

    def render_message(self, record: LogRecord, message: str) -> "ConsoleRenderable":
        """Render message text in to Text.

        Args:
            record (LogRecord): logging Record.
            message (str): String containing log message.

        Returns:
            ConsoleRenderable: Renderable to display log message.
        """
        use_markup = getattr(record, "markup", self.markup)
        message_text = Text.from_markup(message) if use_markup else Text(message)

        highlighter = getattr(record, "highlighter", self.highlighter)
        if highlighter:
            message_text = highlighter(message_text)

        if self.keywords is None:
            self.keywords = self.KEYWORDS

        if self.keywords:
            message_text.highlight_words(self.keywords, "logging.keyword")

        return message_text

    def render(
        self,
        *,
        record: LogRecord,
        traceback: Optional[Traceback],
        message_renderable: "ConsoleRenderable",
    ) -> "ConsoleRenderable":
        """Render log for display.

        Args:
            record (LogRecord): logging Record.
            traceback (Optional[Traceback]): Traceback instance or None for no Traceback.
            message_renderable (ConsoleRenderable): Renderable (typically Text) containing log message contents.

        Returns:
            ConsoleRenderable: Renderable to display log.
        """
        path = Path(record.pathname).name
        level = self.get_level_text(record)
        time_format = None if self.formatter is None else self.formatter.datefmt
        log_time = datetime.fromtimestamp(record.created)

        log_renderable = self._log_render(
            self.console,
            [message_renderable] if not traceback else [message_renderable, traceback],
            log_time=log_time,
            time_format=time_format,
            level=level,
            path=path,
            line_no=record.lineno,
            link_path=record.pathname if self.enable_link_path else None,
        )
        return log_renderable


if __name__ == "__main__":  # pragma: no cover
    from time import sleep

    FORMAT = "%(message)s"
    # FORMAT = "%(asctime)-15s - %(levelname)s - %(message)s"
    logging.basicConfig(
        level="NOTSET",
        format=FORMAT,
        datefmt="[%X]",
        handlers=[RichHandler(rich_tracebacks=True, tracebacks_show_locals=True)],
    )
    log = logging.getLogger("rich")

    log.info("Server starting...")
    log.info("Listening on http://127.0.0.1:8080")
    sleep(1)

    log.info("GET /index.html 200 1298")
    log.info("GET /imgs/backgrounds/back1.jpg 200 54386")
    log.info("GET /css/styles.css 200 54386")
    log.warning("GET /favicon.ico 404 242")
    sleep(1)

    log.debug(
        "JSONRPC request\n--> %r\n<-- %r",
        {
            "version": "1.1",
            "method": "confirmFruitPurchase",
            "params": [["apple", "orange", "mangoes", "pomelo"], 1.123],
            "id": "194521489",
        },
        {"version": "1.1", "result": True, "error": None, "id": "194521489"},
    )
    log.debug(
        "Loading configuration file /adasd/asdasd/qeqwe/qwrqwrqwr/sdgsdgsdg/werwerwer/dfgerert/ertertert/ertetert/werwerwer"
    )
    log.error("Unable to find 'pomelo' in database!")
    log.info("POST /jsonrpc/ 200 65532")
    log.info("POST /admin/ 401 42234")
    log.warning("password was rejected for admin site.")

    def divide() -> None:
        number = 1
        divisor = 0
        foos = ["foo"] * 100
        log.debug("in divide")
        try:
            number / divisor
        except:
            log.exception("An error of some kind occurred!")

    divide()
    sleep(1)
    log.critical("Out of memory!")
    log.info("Server exited with code=-1")
    log.info("[bold]EXITING...[/bold]", extra=dict(markup=True))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\printing\pytorch.py ===

from sympy.printing.pycode import AbstractPythonCodePrinter, ArrayPrinter
from sympy.matrices.expressions import MatrixExpr
from sympy.core.mul import Mul
from sympy.printing.precedence import PRECEDENCE
from sympy.external import import_module
from sympy.codegen.cfunctions import Sqrt
from sympy import S
from sympy import Integer

import sympy

torch = import_module('torch')


class TorchPrinter(ArrayPrinter, AbstractPythonCodePrinter):

    printmethod = "_torchcode"

    mapping = {
        sympy.Abs: "torch.abs",
        sympy.sign: "torch.sign",

        # XXX May raise error for ints.
        sympy.ceiling: "torch.ceil",
        sympy.floor: "torch.floor",
        sympy.log: "torch.log",
        sympy.exp: "torch.exp",
        Sqrt: "torch.sqrt",
        sympy.cos: "torch.cos",
        sympy.acos: "torch.acos",
        sympy.sin: "torch.sin",
        sympy.asin: "torch.asin",
        sympy.tan: "torch.tan",
        sympy.atan: "torch.atan",
        sympy.atan2: "torch.atan2",
        # XXX Also may give NaN for complex results.
        sympy.cosh: "torch.cosh",
        sympy.acosh: "torch.acosh",
        sympy.sinh: "torch.sinh",
        sympy.asinh: "torch.asinh",
        sympy.tanh: "torch.tanh",
        sympy.atanh: "torch.atanh",
        sympy.Pow: "torch.pow",

        sympy.re: "torch.real",
        sympy.im: "torch.imag",
        sympy.arg: "torch.angle",

        # XXX May raise error for ints and complexes
        sympy.erf: "torch.erf",
        sympy.loggamma: "torch.lgamma",

        sympy.Eq: "torch.eq",
        sympy.Ne: "torch.ne",
        sympy.StrictGreaterThan: "torch.gt",
        sympy.StrictLessThan: "torch.lt",
        sympy.LessThan: "torch.le",
        sympy.GreaterThan: "torch.ge",

        sympy.And: "torch.logical_and",
        sympy.Or: "torch.logical_or",
        sympy.Not: "torch.logical_not",
        sympy.Max: "torch.max",
        sympy.Min: "torch.min",

        # Matrices
        sympy.MatAdd: "torch.add",
        sympy.HadamardProduct: "torch.mul",
        sympy.Trace: "torch.trace",

        # XXX May raise error for integer matrices.
        sympy.Determinant: "torch.det",
    }

    _default_settings = dict(
        AbstractPythonCodePrinter._default_settings,
        torch_version=None,
        requires_grad=False,
        dtype="torch.float64",
    )

    def __init__(self, settings=None):
        super().__init__(settings)

        version = self._settings['torch_version']
        self.requires_grad = self._settings['requires_grad']
        self.dtype = self._settings['dtype']
        if version is None and torch:
            version = torch.__version__
        self.torch_version = version

    def _print_Function(self, expr):

        op = self.mapping.get(type(expr), None)
        if op is None:
            return super()._print_Basic(expr)
        children = [self._print(arg) for arg in expr.args]
        if len(children) == 1:
            return "%s(%s)" % (
                self._module_format(op),
                children[0]
            )
        else:
            return self._expand_fold_binary_op(op, children)

    # mirrors the tensorflow version
    _print_Expr = _print_Function
    _print_Application = _print_Function
    _print_MatrixExpr = _print_Function
    _print_Relational = _print_Function
    _print_Not = _print_Function
    _print_And = _print_Function
    _print_Or = _print_Function
    _print_HadamardProduct = _print_Function
    _print_Trace = _print_Function
    _print_Determinant = _print_Function

    def _print_Inverse(self, expr):
        return '{}({})'.format(self._module_format("torch.linalg.inv"),
                               self._print(expr.args[0]))

    def _print_Transpose(self, expr):
        if expr.arg.is_Matrix and expr.arg.shape[0] == expr.arg.shape[1]:
            # For square matrices, we can use the .t() method
            return "{}({}).t()".format("torch.transpose", self._print(expr.arg))
        else:
            # For non-square matrices or more general cases
            # transpose first and second dimensions (typical matrix transpose)
            return "{}.permute({})".format(
                self._print(expr.arg),
                ", ".join([str(i) for i in range(len(expr.arg.shape))])[::-1]
            )

    def _print_PermuteDims(self, expr):
        return "%s.permute(%s)" % (
            self._print(expr.expr),
            ", ".join(str(i) for i in expr.permutation.array_form)
        )

    def _print_Derivative(self, expr):
        # this version handles multi-variable and mixed partial derivatives. The tensorflow version does not.
        variables = expr.variables
        expr_arg = expr.expr

        # Handle multi-variable or repeated derivatives
        if len(variables) > 1 or (
            len(variables) == 1 and not isinstance(variables[0], tuple) and variables.count(variables[0]) > 1):
            result = self._print(expr_arg)
            var_groups = {}

            # Group variables by base symbol
            for var in variables:
                if isinstance(var, tuple):
                    base_var, order = var
                    var_groups[base_var] = var_groups.get(base_var, 0) + order
                else:
                    var_groups[var] = var_groups.get(var, 0) + 1

            # Apply gradients in sequence
            for var, order in var_groups.items():
                for _ in range(order):
                    result = "torch.autograd.grad({}, {}, create_graph=True)[0]".format(result, self._print(var))
            return result

        # Handle single variable case
        if len(variables) == 1:
            variable = variables[0]
            if isinstance(variable, tuple) and len(variable) == 2:
                base_var, order = variable
                if not isinstance(order, Integer): raise NotImplementedError("Only integer orders are supported")
                result = self._print(expr_arg)
                for _ in range(order):
                    result = "torch.autograd.grad({}, {}, create_graph=True)[0]".format(result, self._print(base_var))
                return result
            return "torch.autograd.grad({}, {})[0]".format(self._print(expr_arg), self._print(variable))

        return self._print(expr_arg)  # Empty variables case

    def _print_Piecewise(self, expr):
        from sympy import Piecewise
        e, cond = expr.args[0].args
        if len(expr.args) == 1:
            return '{}({}, {}, {})'.format(
                self._module_format("torch.where"),
                self._print(cond),
                self._print(e),
                0)

        return '{}({}, {}, {})'.format(
            self._module_format("torch.where"),
            self._print(cond),
            self._print(e),
            self._print(Piecewise(*expr.args[1:])))

    def _print_Pow(self, expr):
        # XXX May raise error for
        # int**float or int**complex or float**complex
        base, exp = expr.args
        if expr.exp == S.Half:
            return "{}({})".format(
                self._module_format("torch.sqrt"), self._print(base))
        return "{}({}, {})".format(
            self._module_format("torch.pow"),
            self._print(base), self._print(exp))

    def _print_MatMul(self, expr):
        # Separate matrix and scalar arguments
        mat_args = [arg for arg in expr.args if isinstance(arg, MatrixExpr)]
        args = [arg for arg in expr.args if arg not in mat_args]
        # Handle scalar multipliers if present
        if args:
            return "%s*%s" % (
                self.parenthesize(Mul.fromiter(args), PRECEDENCE["Mul"]),
                self._expand_fold_binary_op("torch.matmul", mat_args)
            )
        else:
            return self._expand_fold_binary_op("torch.matmul", mat_args)

    def _print_MatPow(self, expr):
        return self._expand_fold_binary_op("torch.mm", [expr.base]*expr.exp)

    def _print_MatrixBase(self, expr):
        data = "[" + ", ".join(["[" + ", ".join([self._print(j) for j in i]) + "]" for i in expr.tolist()]) + "]"
        params = [str(data)]
        params.append(f"dtype={self.dtype}")
        if self.requires_grad:
            params.append("requires_grad=True")

        return "{}({})".format(
            self._module_format("torch.tensor"),
            ", ".join(params)
        )

    def _print_isnan(self, expr):
        return f'torch.isnan({self._print(expr.args[0])})'

    def _print_isinf(self, expr):
        return f'torch.isinf({self._print(expr.args[0])})'

    def _print_Identity(self, expr):
        if all(dim.is_Integer for dim in expr.shape):
            return "{}({})".format(
                self._module_format("torch.eye"),
                self._print(expr.shape[0])
            )
        else:
            # For symbolic dimensions, fall back to a more general approach
            return "{}({}, {})".format(
                self._module_format("torch.eye"),
                self._print(expr.shape[0]),
                self._print(expr.shape[1])
            )

    def _print_ZeroMatrix(self, expr):
        return "{}({})".format(
            self._module_format("torch.zeros"),
            self._print(expr.shape)
        )

    def _print_OneMatrix(self, expr):
        return "{}({})".format(
            self._module_format("torch.ones"),
            self._print(expr.shape)
        )

    def _print_conjugate(self, expr):
        return f"{self._module_format('torch.conj')}({self._print(expr.args[0])})"

    def _print_ImaginaryUnit(self, expr):
        return "1j"  # uses the Python built-in 1j notation for the imaginary unit

    def _print_Heaviside(self, expr):
        args = [self._print(expr.args[0]), "0.5"]
        if len(expr.args) > 1:
            args[1] = self._print(expr.args[1])
        return f"{self._module_format('torch.heaviside')}({args[0]}, {args[1]})"

    def _print_gamma(self, expr):
        return f"{self._module_format('torch.special.gamma')}({self._print(expr.args[0])})"

    def _print_polygamma(self, expr):
        if expr.args[0] == S.Zero:
            return f"{self._module_format('torch.special.digamma')}({self._print(expr.args[1])})"
        else:
            raise NotImplementedError("PyTorch only supports digamma (0th order polygamma)")

    _module = "torch"
    _einsum = "einsum"
    _add = "add"
    _transpose = "t"
    _ones = "ones"
    _zeros = "zeros"

def torch_code(expr, requires_grad=False, dtype="torch.float64", **settings):
    printer = TorchPrinter(settings={'requires_grad': requires_grad, 'dtype': dtype})
    return printer.doprint(expr, **settings)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\click\_winconsole.py ===
# This module is based on the excellent work by Adam Bartoš who
# provided a lot of what went into the implementation here in
# the discussion to issue1602 in the Python bug tracker.
#
# There are some general differences in regards to how this works
# compared to the original patches as we do not need to patch
# the entire interpreter but just work in our little world of
# echo and prompt.
from __future__ import annotations

import collections.abc as cabc
import io
import sys
import time
import typing as t
from ctypes import Array
from ctypes import byref
from ctypes import c_char
from ctypes import c_char_p
from ctypes import c_int
from ctypes import c_ssize_t
from ctypes import c_ulong
from ctypes import c_void_p
from ctypes import POINTER
from ctypes import py_object
from ctypes import Structure
from ctypes.wintypes import DWORD
from ctypes.wintypes import HANDLE
from ctypes.wintypes import LPCWSTR
from ctypes.wintypes import LPWSTR

from ._compat import _NonClosingTextIOWrapper

assert sys.platform == "win32"
import msvcrt  # noqa: E402
from ctypes import windll  # noqa: E402
from ctypes import WINFUNCTYPE  # noqa: E402

c_ssize_p = POINTER(c_ssize_t)

kernel32 = windll.kernel32
GetStdHandle = kernel32.GetStdHandle
ReadConsoleW = kernel32.ReadConsoleW
WriteConsoleW = kernel32.WriteConsoleW
GetConsoleMode = kernel32.GetConsoleMode
GetLastError = kernel32.GetLastError
GetCommandLineW = WINFUNCTYPE(LPWSTR)(("GetCommandLineW", windll.kernel32))
CommandLineToArgvW = WINFUNCTYPE(POINTER(LPWSTR), LPCWSTR, POINTER(c_int))(
    ("CommandLineToArgvW", windll.shell32)
)
LocalFree = WINFUNCTYPE(c_void_p, c_void_p)(("LocalFree", windll.kernel32))

STDIN_HANDLE = GetStdHandle(-10)
STDOUT_HANDLE = GetStdHandle(-11)
STDERR_HANDLE = GetStdHandle(-12)

PyBUF_SIMPLE = 0
PyBUF_WRITABLE = 1

ERROR_SUCCESS = 0
ERROR_NOT_ENOUGH_MEMORY = 8
ERROR_OPERATION_ABORTED = 995

STDIN_FILENO = 0
STDOUT_FILENO = 1
STDERR_FILENO = 2

EOF = b"\x1a"
MAX_BYTES_WRITTEN = 32767

if t.TYPE_CHECKING:
    try:
        # Using `typing_extensions.Buffer` instead of `collections.abc`
        # on Windows for some reason does not have `Sized` implemented.
        from collections.abc import Buffer  # type: ignore
    except ImportError:
        from typing_extensions import Buffer

try:
    from ctypes import pythonapi
except ImportError:
    # On PyPy we cannot get buffers so our ability to operate here is
    # severely limited.
    get_buffer = None
else:

    class Py_buffer(Structure):
        _fields_ = [  # noqa: RUF012
            ("buf", c_void_p),
            ("obj", py_object),
            ("len", c_ssize_t),
            ("itemsize", c_ssize_t),
            ("readonly", c_int),
            ("ndim", c_int),
            ("format", c_char_p),
            ("shape", c_ssize_p),
            ("strides", c_ssize_p),
            ("suboffsets", c_ssize_p),
            ("internal", c_void_p),
        ]

    PyObject_GetBuffer = pythonapi.PyObject_GetBuffer
    PyBuffer_Release = pythonapi.PyBuffer_Release

    def get_buffer(obj: Buffer, writable: bool = False) -> Array[c_char]:
        buf = Py_buffer()
        flags: int = PyBUF_WRITABLE if writable else PyBUF_SIMPLE
        PyObject_GetBuffer(py_object(obj), byref(buf), flags)

        try:
            buffer_type = c_char * buf.len
            out: Array[c_char] = buffer_type.from_address(buf.buf)
            return out
        finally:
            PyBuffer_Release(byref(buf))


class _WindowsConsoleRawIOBase(io.RawIOBase):
    def __init__(self, handle: int | None) -> None:
        self.handle = handle

    def isatty(self) -> t.Literal[True]:
        super().isatty()
        return True


class _WindowsConsoleReader(_WindowsConsoleRawIOBase):
    def readable(self) -> t.Literal[True]:
        return True

    def readinto(self, b: Buffer) -> int:
        bytes_to_be_read = len(b)
        if not bytes_to_be_read:
            return 0
        elif bytes_to_be_read % 2:
            raise ValueError(
                "cannot read odd number of bytes from UTF-16-LE encoded console"
            )

        buffer = get_buffer(b, writable=True)
        code_units_to_be_read = bytes_to_be_read // 2
        code_units_read = c_ulong()

        rv = ReadConsoleW(
            HANDLE(self.handle),
            buffer,
            code_units_to_be_read,
            byref(code_units_read),
            None,
        )
        if GetLastError() == ERROR_OPERATION_ABORTED:
            # wait for KeyboardInterrupt
            time.sleep(0.1)
        if not rv:
            raise OSError(f"Windows error: {GetLastError()}")

        if buffer[0] == EOF:
            return 0
        return 2 * code_units_read.value


class _WindowsConsoleWriter(_WindowsConsoleRawIOBase):
    def writable(self) -> t.Literal[True]:
        return True

    @staticmethod
    def _get_error_message(errno: int) -> str:
        if errno == ERROR_SUCCESS:
            return "ERROR_SUCCESS"
        elif errno == ERROR_NOT_ENOUGH_MEMORY:
            return "ERROR_NOT_ENOUGH_MEMORY"
        return f"Windows error {errno}"

    def write(self, b: Buffer) -> int:
        bytes_to_be_written = len(b)
        buf = get_buffer(b)
        code_units_to_be_written = min(bytes_to_be_written, MAX_BYTES_WRITTEN) // 2
        code_units_written = c_ulong()

        WriteConsoleW(
            HANDLE(self.handle),
            buf,
            code_units_to_be_written,
            byref(code_units_written),
            None,
        )
        bytes_written = 2 * code_units_written.value

        if bytes_written == 0 and bytes_to_be_written > 0:
            raise OSError(self._get_error_message(GetLastError()))
        return bytes_written


class ConsoleStream:
    def __init__(self, text_stream: t.TextIO, byte_stream: t.BinaryIO) -> None:
        self._text_stream = text_stream
        self.buffer = byte_stream

    @property
    def name(self) -> str:
        return self.buffer.name

    def write(self, x: t.AnyStr) -> int:
        if isinstance(x, str):
            return self._text_stream.write(x)
        try:
            self.flush()
        except Exception:
            pass
        return self.buffer.write(x)

    def writelines(self, lines: cabc.Iterable[t.AnyStr]) -> None:
        for line in lines:
            self.write(line)

    def __getattr__(self, name: str) -> t.Any:
        return getattr(self._text_stream, name)

    def isatty(self) -> bool:
        return self.buffer.isatty()

    def __repr__(self) -> str:
        return f"<ConsoleStream name={self.name!r} encoding={self.encoding!r}>"


def _get_text_stdin(buffer_stream: t.BinaryIO) -> t.TextIO:
    text_stream = _NonClosingTextIOWrapper(
        io.BufferedReader(_WindowsConsoleReader(STDIN_HANDLE)),
        "utf-16-le",
        "strict",
        line_buffering=True,
    )
    return t.cast(t.TextIO, ConsoleStream(text_stream, buffer_stream))


def _get_text_stdout(buffer_stream: t.BinaryIO) -> t.TextIO:
    text_stream = _NonClosingTextIOWrapper(
        io.BufferedWriter(_WindowsConsoleWriter(STDOUT_HANDLE)),
        "utf-16-le",
        "strict",
        line_buffering=True,
    )
    return t.cast(t.TextIO, ConsoleStream(text_stream, buffer_stream))


def _get_text_stderr(buffer_stream: t.BinaryIO) -> t.TextIO:
    text_stream = _NonClosingTextIOWrapper(
        io.BufferedWriter(_WindowsConsoleWriter(STDERR_HANDLE)),
        "utf-16-le",
        "strict",
        line_buffering=True,
    )
    return t.cast(t.TextIO, ConsoleStream(text_stream, buffer_stream))


_stream_factories: cabc.Mapping[int, t.Callable[[t.BinaryIO], t.TextIO]] = {
    0: _get_text_stdin,
    1: _get_text_stdout,
    2: _get_text_stderr,
}


def _is_console(f: t.TextIO) -> bool:
    if not hasattr(f, "fileno"):
        return False

    try:
        fileno = f.fileno()
    except (OSError, io.UnsupportedOperation):
        return False

    handle = msvcrt.get_osfhandle(fileno)
    return bool(GetConsoleMode(handle, byref(DWORD())))


def _get_windows_console_stream(
    f: t.TextIO, encoding: str | None, errors: str | None
) -> t.TextIO | None:
    if (
        get_buffer is None
        or encoding not in {"utf-16-le", None}
        or errors not in {"strict", None}
        or not _is_console(f)
    ):
        return None

    func = _stream_factories.get(f.fileno())
    if func is None:
        return None

    b = getattr(f, "buffer", None)

    if b is None:
        return None

    return func(b)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\utils\temp_dir.py ===
import errno
import itertools
import logging
import os.path
import tempfile
import traceback
from contextlib import ExitStack, contextmanager
from pathlib import Path
from typing import (
    Any,
    Callable,
    Dict,
    Generator,
    List,
    Optional,
    TypeVar,
    Union,
)

from pip._internal.utils.misc import enum, rmtree

logger = logging.getLogger(__name__)

_T = TypeVar("_T", bound="TempDirectory")


# Kinds of temporary directories. Only needed for ones that are
# globally-managed.
tempdir_kinds = enum(
    BUILD_ENV="build-env",
    EPHEM_WHEEL_CACHE="ephem-wheel-cache",
    REQ_BUILD="req-build",
)


_tempdir_manager: Optional[ExitStack] = None


@contextmanager
def global_tempdir_manager() -> Generator[None, None, None]:
    global _tempdir_manager
    with ExitStack() as stack:
        old_tempdir_manager, _tempdir_manager = _tempdir_manager, stack
        try:
            yield
        finally:
            _tempdir_manager = old_tempdir_manager


class TempDirectoryTypeRegistry:
    """Manages temp directory behavior"""

    def __init__(self) -> None:
        self._should_delete: Dict[str, bool] = {}

    def set_delete(self, kind: str, value: bool) -> None:
        """Indicate whether a TempDirectory of the given kind should be
        auto-deleted.
        """
        self._should_delete[kind] = value

    def get_delete(self, kind: str) -> bool:
        """Get configured auto-delete flag for a given TempDirectory type,
        default True.
        """
        return self._should_delete.get(kind, True)


_tempdir_registry: Optional[TempDirectoryTypeRegistry] = None


@contextmanager
def tempdir_registry() -> Generator[TempDirectoryTypeRegistry, None, None]:
    """Provides a scoped global tempdir registry that can be used to dictate
    whether directories should be deleted.
    """
    global _tempdir_registry
    old_tempdir_registry = _tempdir_registry
    _tempdir_registry = TempDirectoryTypeRegistry()
    try:
        yield _tempdir_registry
    finally:
        _tempdir_registry = old_tempdir_registry


class _Default:
    pass


_default = _Default()


class TempDirectory:
    """Helper class that owns and cleans up a temporary directory.

    This class can be used as a context manager or as an OO representation of a
    temporary directory.

    Attributes:
        path
            Location to the created temporary directory
        delete
            Whether the directory should be deleted when exiting
            (when used as a contextmanager)

    Methods:
        cleanup()
            Deletes the temporary directory

    When used as a context manager, if the delete attribute is True, on
    exiting the context the temporary directory is deleted.
    """

    def __init__(
        self,
        path: Optional[str] = None,
        delete: Union[bool, None, _Default] = _default,
        kind: str = "temp",
        globally_managed: bool = False,
        ignore_cleanup_errors: bool = True,
    ):
        super().__init__()

        if delete is _default:
            if path is not None:
                # If we were given an explicit directory, resolve delete option
                # now.
                delete = False
            else:
                # Otherwise, we wait until cleanup and see what
                # tempdir_registry says.
                delete = None

        # The only time we specify path is in for editables where it
        # is the value of the --src option.
        if path is None:
            path = self._create(kind)

        self._path = path
        self._deleted = False
        self.delete = delete
        self.kind = kind
        self.ignore_cleanup_errors = ignore_cleanup_errors

        if globally_managed:
            assert _tempdir_manager is not None
            _tempdir_manager.enter_context(self)

    @property
    def path(self) -> str:
        assert not self._deleted, f"Attempted to access deleted path: {self._path}"
        return self._path

    def __repr__(self) -> str:
        return f"<{self.__class__.__name__} {self.path!r}>"

    def __enter__(self: _T) -> _T:
        return self

    def __exit__(self, exc: Any, value: Any, tb: Any) -> None:
        if self.delete is not None:
            delete = self.delete
        elif _tempdir_registry:
            delete = _tempdir_registry.get_delete(self.kind)
        else:
            delete = True

        if delete:
            self.cleanup()

    def _create(self, kind: str) -> str:
        """Create a temporary directory and store its path in self.path"""
        # We realpath here because some systems have their default tmpdir
        # symlinked to another directory.  This tends to confuse build
        # scripts, so we canonicalize the path by traversing potential
        # symlinks here.
        path = os.path.realpath(tempfile.mkdtemp(prefix=f"pip-{kind}-"))
        logger.debug("Created temporary directory: %s", path)
        return path

    def cleanup(self) -> None:
        """Remove the temporary directory created and reset state"""
        self._deleted = True
        if not os.path.exists(self._path):
            return

        errors: List[BaseException] = []

        def onerror(
            func: Callable[..., Any],
            path: Path,
            exc_val: BaseException,
        ) -> None:
            """Log a warning for a `rmtree` error and continue"""
            formatted_exc = "\n".join(
                traceback.format_exception_only(type(exc_val), exc_val)
            )
            formatted_exc = formatted_exc.rstrip()  # remove trailing new line
            if func in (os.unlink, os.remove, os.rmdir):
                logger.debug(
                    "Failed to remove a temporary file '%s' due to %s.\n",
                    path,
                    formatted_exc,
                )
            else:
                logger.debug("%s failed with %s.", func.__qualname__, formatted_exc)
            errors.append(exc_val)

        if self.ignore_cleanup_errors:
            try:
                # first try with @retry; retrying to handle ephemeral errors
                rmtree(self._path, ignore_errors=False)
            except OSError:
                # last pass ignore/log all errors
                rmtree(self._path, onexc=onerror)
            if errors:
                logger.warning(
                    "Failed to remove contents in a temporary directory '%s'.\n"
                    "You can safely remove it manually.",
                    self._path,
                )
        else:
            rmtree(self._path)


class AdjacentTempDirectory(TempDirectory):
    """Helper class that creates a temporary directory adjacent to a real one.

    Attributes:
        original
            The original directory to create a temp directory for.
        path
            After calling create() or entering, contains the full
            path to the temporary directory.
        delete
            Whether the directory should be deleted when exiting
            (when used as a contextmanager)

    """

    # The characters that may be used to name the temp directory
    # We always prepend a ~ and then rotate through these until
    # a usable name is found.
    # pkg_resources raises a different error for .dist-info folder
    # with leading '-' and invalid metadata
    LEADING_CHARS = "-~.=%0123456789"

    def __init__(self, original: str, delete: Optional[bool] = None) -> None:
        self.original = original.rstrip("/\\")
        super().__init__(delete=delete)

    @classmethod
    def _generate_names(cls, name: str) -> Generator[str, None, None]:
        """Generates a series of temporary names.

        The algorithm replaces the leading characters in the name
        with ones that are valid filesystem characters, but are not
        valid package names (for both Python and pip definitions of
        package).
        """
        for i in range(1, len(name)):
            for candidate in itertools.combinations_with_replacement(
                cls.LEADING_CHARS, i - 1
            ):
                new_name = "~" + "".join(candidate) + name[i:]
                if new_name != name:
                    yield new_name

        # If we make it this far, we will have to make a longer name
        for i in range(len(cls.LEADING_CHARS)):
            for candidate in itertools.combinations_with_replacement(
                cls.LEADING_CHARS, i
            ):
                new_name = "~" + "".join(candidate) + name
                if new_name != name:
                    yield new_name

    def _create(self, kind: str) -> str:
        root, name = os.path.split(self.original)
        for candidate in self._generate_names(name):
            path = os.path.join(root, candidate)
            try:
                os.mkdir(path)
            except OSError as ex:
                # Continue if the name exists already
                if ex.errno != errno.EEXIST:
                    raise
            else:
                path = os.path.realpath(path)
                break
        else:
            # Final fallback on the default behavior.
            path = os.path.realpath(tempfile.mkdtemp(prefix=f"pip-{kind}-"))

        logger.debug("Created temporary directory: %s", path)
        return path

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\stable_diffusion\engine_builder.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
import hashlib
import os
from enum import Enum

import torch
from diffusion_models import CLIP, VAE, CLIPWithProj, PipelineInfo, UNet, UNetXL


class EngineType(Enum):
    ORT_CUDA = 0  # ONNX Runtime CUDA Execution Provider
    ORT_TRT = 1  # ONNX Runtime TensorRT Execution Provider
    TRT = 2  # TensorRT
    TORCH = 3  # PyTorch


def get_engine_type(name: str) -> EngineType:
    name_to_type = {
        "ORT_CUDA": EngineType.ORT_CUDA,
        "ORT_TRT": EngineType.ORT_TRT,
        "TRT": EngineType.TRT,
        "TORCH": EngineType.TORCH,
    }
    return name_to_type[name]


class EngineBuilder:
    def __init__(
        self,
        engine_type: EngineType,
        pipeline_info: PipelineInfo,
        device="cuda",
        max_batch_size=16,
        use_cuda_graph=False,
    ):
        """
        Initializes the Engine Builder.

        Args:
            pipeline_info (PipelineInfo):
                Version and Type of pipeline.
            device (str | torch.device):
                device to run engine
            max_batch_size (int):
                Maximum batch size for dynamic batch engine.
            use_cuda_graph (bool):
                Use CUDA graph to capture engine execution and then launch inference
        """
        self.engine_type = engine_type
        self.pipeline_info = pipeline_info
        self.max_batch_size = max_batch_size
        self.use_cuda_graph = use_cuda_graph
        self.device = torch.device(device)
        self.torch_device = torch.device(device, torch.cuda.current_device())
        self.stages = pipeline_info.stages()

        self.vae_torch_fallback = self.pipeline_info.vae_torch_fallback() and self.engine_type != EngineType.TORCH
        self.custom_fp16_vae = self.pipeline_info.custom_fp16_vae()

        self.models = {}
        self.engines = {}
        self.torch_models = {}
        self.use_vae_slicing = False

        self.torch_sdpa = getattr(torch.nn.functional, "scaled_dot_product_attention", None)

    def enable_vae_slicing(self):
        self.use_vae_slicing = True

    def disable_torch_spda(self):
        if hasattr(torch.nn.functional, "scaled_dot_product_attention"):
            delattr(torch.nn.functional, "scaled_dot_product_attention")

    def enable_torch_spda(self):
        if (not hasattr(torch.nn.functional, "scaled_dot_product_attention")) and self.torch_sdpa:
            torch.nn.functional.scaled_dot_product_attention = self.torch_sdpa

    def teardown(self):
        for engine in self.engines.values():
            del engine
        self.engines = {}

    def get_diffusers_module_name(self, model_name):
        name_mapping = {
            "clip": "text_encoder",
            "clip2": "text_encoder_2",
            "unet": "unet",
            "unetxl": "unet",
            "vae": "vae_decoder",
        }
        return name_mapping.get(model_name, model_name)

    def get_cached_model_name(self, model_name):
        model_name = self.get_diffusers_module_name(model_name)
        is_unet = model_name == "unet"
        hash_source = []
        if model_name in ["text_encoder", "text_encoder_2", "unet"] and self.pipeline_info.lora_weights:
            if self.pipeline_info.lora_weights in [
                "latent-consistency/lcm-lora-sdxl",
                "latent-consistency/lcm-lora-sdv1-5",
            ]:
                if is_unet:
                    model_name = "unet_lcm-lora"
            else:
                model_name = model_name + "_lora"
                hash_source.append(self.pipeline_info.lora_weights)

        # TODO(tianleiwu): save custom model to a directory named by its original model.
        if is_unet and self.pipeline_info.custom_unet():
            model_name = model_name + "_lcm"

        if model_name in ["unet"] and self.pipeline_info.controlnet:
            model_name = model_name + "_" + "_".join(self.pipeline_info.controlnet)

        if hash_source:
            model_name += "_" + hashlib.sha256("\t".join(hash_source).encode("utf-8")).hexdigest()[:8]

        # TODO: When we support original VAE, we shall save custom VAE to another directory.

        if self.pipeline_info.is_inpaint():
            model_name += "_inpaint"
        return model_name

    def get_model_dir(self, model_name, root_dir, opt=True, suffix="", create=True):
        engine_name = self.engine_type.name.lower()
        if engine_name != "ort_cuda" and not suffix:
            suffix = f".{engine_name}" if opt else ""
        directory_name = self.get_cached_model_name(model_name) + suffix
        onnx_model_dir = os.path.join(root_dir, directory_name)
        if create:
            os.makedirs(onnx_model_dir, exist_ok=True)
        return onnx_model_dir

    def get_onnx_path(self, model_name, onnx_dir, opt=True, suffix=""):
        onnx_model_dir = self.get_model_dir(model_name, onnx_dir, opt=opt, suffix=suffix)
        return os.path.join(onnx_model_dir, "model.onnx")

    def get_engine_path(self, engine_dir, model_name, profile_id):
        return os.path.join(engine_dir, self.get_cached_model_name(model_name) + profile_id)

    def load_pipeline_with_lora(self):
        """Load text encoders and UNet with diffusers pipeline"""
        from diffusers import DiffusionPipeline

        pipeline = DiffusionPipeline.from_pretrained(
            self.pipeline_info.name(),
            variant="fp16",
            torch_dtype=torch.float16,
        )
        pipeline.load_lora_weights(self.pipeline_info.lora_weights)
        pipeline.fuse_lora(lora_scale=self.pipeline_info.lora_scale)

        del pipeline.vae
        pipeline.vae = None
        return pipeline

    def get_or_load_model(self, pipeline, model_name, model_obj, framework_model_dir):
        if model_name in ["clip", "clip2", "unet", "unetxl"] and pipeline:
            if model_name == "clip":
                model = pipeline.text_encoder
                pipeline.text_encoder = None
            elif model_name == "clip2":
                model = pipeline.text_encoder_2
                pipeline.text_encoder_2 = None
            else:
                model = pipeline.unet
                pipeline.unet = None
        else:
            model = model_obj.load_model(framework_model_dir)

        return model.to(self.torch_device)

    def load_models(self, framework_model_dir: str):
        # For TRT or ORT_TRT, we will export fp16 torch model for UNet and VAE
        # For ORT_CUDA, we export fp32 model first, then optimize to fp16.
        export_fp16 = self.engine_type in [EngineType.ORT_TRT, EngineType.TRT]

        if "clip" in self.stages:
            self.models["clip"] = CLIP(
                self.pipeline_info,
                None,  # not loaded yet
                device=self.torch_device,
                max_batch_size=self.max_batch_size,
                clip_skip=0,
            )

        if "clip2" in self.stages:
            self.models["clip2"] = CLIPWithProj(
                self.pipeline_info,
                None,  # not loaded yet
                device=self.torch_device,
                max_batch_size=self.max_batch_size,
                clip_skip=0,
            )

        if "unet" in self.stages:
            self.models["unet"] = UNet(
                self.pipeline_info,
                None,  # not loaded yet
                device=self.torch_device,
                fp16=export_fp16,
                max_batch_size=self.max_batch_size,
                unet_dim=(9 if self.pipeline_info.is_inpaint() else 4),
            )

        if "unetxl" in self.stages:
            self.models["unetxl"] = UNetXL(
                self.pipeline_info,
                None,  # not loaded yet
                device=self.torch_device,
                fp16=export_fp16,
                max_batch_size=self.max_batch_size,
                unet_dim=4,
                time_dim=(5 if self.pipeline_info.is_xl_refiner() else 6),
            )

        # VAE Decoder
        if "vae" in self.stages:
            self.models["vae"] = VAE(
                self.pipeline_info,
                None,  # not loaded yet
                device=self.torch_device,
                max_batch_size=self.max_batch_size,
                fp16=export_fp16,
                custom_fp16_vae=self.custom_fp16_vae,
            )

            if self.vae_torch_fallback:
                self.torch_models["vae"] = self.models["vae"].load_model(framework_model_dir)

    def load_resources(self, image_height, image_width, batch_size):
        if self.engine_type == EngineType.TORCH:
            return

        # Allocate buffers for I/O bindings
        for model_name, obj in self.models.items():
            if model_name == "vae" and self.vae_torch_fallback:
                continue
            slice_size = 1 if (model_name == "vae" and self.use_vae_slicing) else batch_size
            self.engines[model_name].allocate_buffers(
                shape_dict=obj.get_shape_dict(slice_size, image_height, image_width), device=self.torch_device
            )

    def _vae_decode(self, latents):
        if self.engine_type == EngineType.TORCH:
            if self.pipeline_info.is_xl() and not self.custom_fp16_vae:  # need upcast
                latents = latents.to(dtype=torch.float32)
                images = self.engines["vae"](latents)["sample"]
            else:
                images = self.engines["vae"](latents)["sample"]
        elif self.vae_torch_fallback:
            if not self.custom_fp16_vae:
                latents = latents.to(dtype=torch.float32)
                self.torch_models["vae"] = self.torch_models["vae"].to(dtype=torch.float32)
            images = self.torch_models["vae"](latents)["sample"]
        else:
            if self.pipeline_info.is_xl() and not self.custom_fp16_vae:  # need upcast
                images = self.run_engine("vae", {"latent": latents.to(dtype=torch.float32)})["images"]
            else:
                images = self.run_engine("vae", {"latent": latents})["images"]

        return images

    def vae_decode(self, latents):
        if self.use_vae_slicing:
            # The output tensor points to same buffer. Need clone it to avoid overwritten.
            decoded_slices = [self._vae_decode(z_slice).clone() for z_slice in latents.split(1)]
            return torch.cat(decoded_slices)

        return self._vae_decode(latents)


def get_engine_paths(
    work_dir: str, pipeline_info: PipelineInfo, engine_type: EngineType, framework_model_dir: str | None = None
):
    root_dir = work_dir or "."
    short_name = pipeline_info.short_name()

    # When both ORT_CUDA and ORT_TRT/TRT is used, we shall make sub directory for each engine since
    # ORT_CUDA need fp32 torch model, while ORT_TRT/TRT use fp16 torch model.
    onnx_dir = os.path.join(root_dir, engine_type.name, short_name, "onnx")
    engine_dir = os.path.join(root_dir, engine_type.name, short_name, "engine")
    output_dir = os.path.join(root_dir, engine_type.name, short_name, "output")

    timing_cache = os.path.join(root_dir, engine_type.name, "timing_cache")

    # Shared among ORT_CUDA, ORT_TRT and TRT engines, and need use load_model(..., always_download_fp16=True)
    # So that the shared model is always fp16.
    if framework_model_dir is None:
        framework_model_dir = os.path.join(root_dir, "torch_model")

    return onnx_dir, engine_dir, output_dir, framework_model_dir, timing_cache