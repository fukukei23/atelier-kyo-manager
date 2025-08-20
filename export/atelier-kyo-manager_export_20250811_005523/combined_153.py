
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\_numba\kernels\mean_.py ===
"""
Numba 1D mean kernels that can be shared by
* Dataframe / Series
* groupby
* rolling / expanding

Mirrors pandas/_libs/window/aggregation.pyx
"""
from __future__ import annotations

from typing import TYPE_CHECKING

import numba
import numpy as np

from pandas.core._numba.kernels.shared import is_monotonic_increasing
from pandas.core._numba.kernels.sum_ import grouped_kahan_sum

if TYPE_CHECKING:
    from pandas._typing import npt


@numba.jit(nopython=True, nogil=True, parallel=False)
def add_mean(
    val: float,
    nobs: int,
    sum_x: float,
    neg_ct: int,
    compensation: float,
    num_consecutive_same_value: int,
    prev_value: float,
) -> tuple[int, float, int, float, int, float]:
    if not np.isnan(val):
        nobs += 1
        y = val - compensation
        t = sum_x + y
        compensation = t - sum_x - y
        sum_x = t
        if val < 0:
            neg_ct += 1

        if val == prev_value:
            num_consecutive_same_value += 1
        else:
            num_consecutive_same_value = 1
        prev_value = val

    return nobs, sum_x, neg_ct, compensation, num_consecutive_same_value, prev_value


@numba.jit(nopython=True, nogil=True, parallel=False)
def remove_mean(
    val: float, nobs: int, sum_x: float, neg_ct: int, compensation: float
) -> tuple[int, float, int, float]:
    if not np.isnan(val):
        nobs -= 1
        y = -val - compensation
        t = sum_x + y
        compensation = t - sum_x - y
        sum_x = t
        if val < 0:
            neg_ct -= 1
    return nobs, sum_x, neg_ct, compensation


@numba.jit(nopython=True, nogil=True, parallel=False)
def sliding_mean(
    values: np.ndarray,
    result_dtype: np.dtype,
    start: np.ndarray,
    end: np.ndarray,
    min_periods: int,
) -> tuple[np.ndarray, list[int]]:
    N = len(start)
    nobs = 0
    sum_x = 0.0
    neg_ct = 0
    compensation_add = 0.0
    compensation_remove = 0.0

    is_monotonic_increasing_bounds = is_monotonic_increasing(
        start
    ) and is_monotonic_increasing(end)

    output = np.empty(N, dtype=result_dtype)

    for i in range(N):
        s = start[i]
        e = end[i]
        if i == 0 or not is_monotonic_increasing_bounds:
            prev_value = values[s]
            num_consecutive_same_value = 0

            for j in range(s, e):
                val = values[j]
                (
                    nobs,
                    sum_x,
                    neg_ct,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,
                ) = add_mean(
                    val,
                    nobs,
                    sum_x,
                    neg_ct,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,  # pyright: ignore[reportGeneralTypeIssues]
                )
        else:
            for j in range(start[i - 1], s):
                val = values[j]
                nobs, sum_x, neg_ct, compensation_remove = remove_mean(
                    val, nobs, sum_x, neg_ct, compensation_remove
                )

            for j in range(end[i - 1], e):
                val = values[j]
                (
                    nobs,
                    sum_x,
                    neg_ct,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,
                ) = add_mean(
                    val,
                    nobs,
                    sum_x,
                    neg_ct,
                    compensation_add,
                    num_consecutive_same_value,
                    prev_value,  # pyright: ignore[reportGeneralTypeIssues]
                )

        if nobs >= min_periods and nobs > 0:
            result = sum_x / nobs
            if num_consecutive_same_value >= nobs:
                result = prev_value
            elif neg_ct == 0 and result < 0:
                result = 0
            elif neg_ct == nobs and result > 0:
                result = 0
        else:
            result = np.nan

        output[i] = result

        if not is_monotonic_increasing_bounds:
            nobs = 0
            sum_x = 0.0
            neg_ct = 0
            compensation_remove = 0.0

    # na_position is empty list since float64 can already hold nans
    # Do list comprehension, since numba cannot figure out that na_pos is
    # empty list of ints on its own
    na_pos = [0 for i in range(0)]
    return output, na_pos


@numba.jit(nopython=True, nogil=True, parallel=False)
def grouped_mean(
    values: np.ndarray,
    result_dtype: np.dtype,
    labels: npt.NDArray[np.intp],
    ngroups: int,
    min_periods: int,
) -> tuple[np.ndarray, list[int]]:
    output, nobs_arr, comp_arr, consecutive_counts, prev_vals = grouped_kahan_sum(
        values, result_dtype, labels, ngroups
    )

    # Post-processing, replace sums that don't satisfy min_periods
    for lab in range(ngroups):
        nobs = nobs_arr[lab]
        num_consecutive_same_value = consecutive_counts[lab]
        prev_value = prev_vals[lab]
        sum_x = output[lab]
        if nobs >= min_periods:
            if num_consecutive_same_value >= nobs:
                result = prev_value * nobs
            else:
                result = sum_x
        else:
            result = np.nan
        result /= nobs
        output[lab] = result

    # na_position is empty list since float64 can already hold nans
    # Do list comprehension, since numba cannot figure out that na_pos is
    # empty list of ints on its own
    na_pos = [0 for i in range(0)]
    return output, na_pos

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\resolvelib\providers.py ===
from __future__ import annotations

from typing import (
    TYPE_CHECKING,
    Generic,
    Iterable,
    Iterator,
    Mapping,
    Sequence,
)

from .structs import CT, KT, RT, Matches, RequirementInformation

if TYPE_CHECKING:
    from typing import Any, Protocol

    class Preference(Protocol):
        def __lt__(self, __other: Any) -> bool: ...


class AbstractProvider(Generic[RT, CT, KT]):
    """Delegate class to provide the required interface for the resolver."""

    def identify(self, requirement_or_candidate: RT | CT) -> KT:
        """Given a requirement or candidate, return an identifier for it.

        This is used to identify, e.g. whether two requirements
        should have their specifier parts merged or a candidate matches a
        requirement via ``find_matches()``.
        """
        raise NotImplementedError

    def get_preference(
        self,
        identifier: KT,
        resolutions: Mapping[KT, CT],
        candidates: Mapping[KT, Iterator[CT]],
        information: Mapping[KT, Iterator[RequirementInformation[RT, CT]]],
        backtrack_causes: Sequence[RequirementInformation[RT, CT]],
    ) -> Preference:
        """Produce a sort key for given requirement based on preference.

        As this is a sort key it will be called O(n) times per backtrack
        step, where n is the number of `identifier`s, if you have a check
        which is expensive in some sense. E.g. It needs to make O(n) checks
        per call or takes significant wall clock time, consider using
        `narrow_requirement_selection` to filter the `identifier`s, which
        is applied before this sort key is called.

        The preference is defined as "I think this requirement should be
        resolved first". The lower the return value is, the more preferred
        this group of arguments is.

        :param identifier: An identifier as returned by ``identify()``. This
            identifies the requirement being considered.
        :param resolutions: Mapping of candidates currently pinned by the
            resolver. Each key is an identifier, and the value is a candidate.
            The candidate may conflict with requirements from ``information``.
        :param candidates: Mapping of each dependency's possible candidates.
            Each value is an iterator of candidates.
        :param information: Mapping of requirement information of each package.
            Each value is an iterator of *requirement information*.
        :param backtrack_causes: Sequence of *requirement information* that are
            the requirements that caused the resolver to most recently
            backtrack.

        A *requirement information* instance is a named tuple with two members:

        * ``requirement`` specifies a requirement contributing to the current
          list of candidates.
        * ``parent`` specifies the candidate that provides (depended on) the
          requirement, or ``None`` to indicate a root requirement.

        The preference could depend on various issues, including (not
        necessarily in this order):

        * Is this package pinned in the current resolution result?
        * How relaxed is the requirement? Stricter ones should probably be
          worked on first? (I don't know, actually.)
        * How many possibilities are there to satisfy this requirement? Those
          with few left should likely be worked on first, I guess?
        * Are there any known conflicts for this requirement? We should
          probably work on those with the most known conflicts.

        A sortable value should be returned (this will be used as the ``key``
        parameter of the built-in sorting function). The smaller the value is,
        the more preferred this requirement is (i.e. the sorting function
        is called with ``reverse=False``).
        """
        raise NotImplementedError

    def find_matches(
        self,
        identifier: KT,
        requirements: Mapping[KT, Iterator[RT]],
        incompatibilities: Mapping[KT, Iterator[CT]],
    ) -> Matches[CT]:
        """Find all possible candidates that satisfy the given constraints.

        :param identifier: An identifier as returned by ``identify()``. All
            candidates returned by this method should produce the same
            identifier.
        :param requirements: A mapping of requirements that all returned
            candidates must satisfy. Each key is an identifier, and the value
            an iterator of requirements for that dependency.
        :param incompatibilities: A mapping of known incompatibile candidates of
            each dependency. Each key is an identifier, and the value an
            iterator of incompatibilities known to the resolver. All
            incompatibilities *must* be excluded from the return value.

        This should try to get candidates based on the requirements' types.
        For VCS, local, and archive requirements, the one-and-only match is
        returned, and for a "named" requirement, the index(es) should be
        consulted to find concrete candidates for this requirement.

        The return value should produce candidates ordered by preference; the
        most preferred candidate should come first. The return type may be one
        of the following:

        * A callable that returns an iterator that yields candidates.
        * An collection of candidates.
        * An iterable of candidates. This will be consumed immediately into a
          list of candidates.
        """
        raise NotImplementedError

    def is_satisfied_by(self, requirement: RT, candidate: CT) -> bool:
        """Whether the given requirement can be satisfied by a candidate.

        The candidate is guaranteed to have been generated from the
        requirement.

        A boolean should be returned to indicate whether ``candidate`` is a
        viable solution to the requirement.
        """
        raise NotImplementedError

    def get_dependencies(self, candidate: CT) -> Iterable[RT]:
        """Get dependencies of a candidate.

        This should return a collection of requirements that `candidate`
        specifies as its dependencies.
        """
        raise NotImplementedError

    def narrow_requirement_selection(
        self,
        identifiers: Iterable[KT],
        resolutions: Mapping[KT, CT],
        candidates: Mapping[KT, Iterator[CT]],
        information: Mapping[KT, Iterator[RequirementInformation[RT, CT]]],
        backtrack_causes: Sequence[RequirementInformation[RT, CT]],
    ) -> Iterable[KT]:
        """
        An optional method to narrow the selection of requirements being
        considered during resolution. This method is called O(1) time per
        backtrack step.

        :param identifiers: An iterable of `identifiers` as returned by
            ``identify()``. These identify all requirements currently being
            considered.
        :param resolutions: A mapping of candidates currently pinned by the
            resolver. Each key is an identifier, and the value is a candidate
            that may conflict with requirements from ``information``.
        :param candidates: A mapping of each dependency's possible candidates.
            Each value is an iterator of candidates.
        :param information: A mapping of requirement information for each package.
            Each value is an iterator of *requirement information*.
        :param backtrack_causes: A sequence of *requirement information* that are
            the requirements causing the resolver to most recently
            backtrack.

        A *requirement information* instance is a named tuple with two members:

        * ``requirement`` specifies a requirement contributing to the current
          list of candidates.
        * ``parent`` specifies the candidate that provides (is depended on for)
          the requirement, or ``None`` to indicate a root requirement.

        Must return a non-empty subset of `identifiers`, with the default
        implementation being to return `identifiers` unchanged. Those `identifiers`
        will then be passed to the sort key `get_preference` to pick the most
        prefered requirement to attempt to pin, unless `narrow_requirement_selection`
        returns only 1 requirement, in which case that will be used without
        calling the sort key `get_preference`.

        This method is designed to be used by the provider to optimize the
        dependency resolution, e.g. if a check cost is O(m) and it can be done
        against all identifiers at once then filtering the requirement selection
        here will cost O(m) but making it part of the sort key in `get_preference`
        will cost O(m*n), where n is the number of `identifiers`.

        Returns:
            Iterable[KT]: A non-empty subset of `identifiers`.
        """
        return identifiers

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v135\bluetooth_emulation.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: BluetoothEmulation (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

class CentralState(enum.Enum):
    '''
    Indicates the various states of Central.
    '''
    ABSENT = "absent"
    POWERED_OFF = "powered-off"
    POWERED_ON = "powered-on"

    def to_json(self):
        return self.value

    @classmethod
    def from_json(cls, json):
        return cls(json)


@dataclass
class ManufacturerData:
    '''
    Stores the manufacturer data
    '''
    #: Company identifier
    #: https://bitbucket.org/bluetooth-SIG/public/src/main/assigned_numbers/company_identifiers/company_identifiers.yaml
    #: https://usb.org/developers
    key: int

    #: Manufacturer-specific data
    data: str

    def to_json(self):
        json = dict()
        json['key'] = self.key
        json['data'] = self.data
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            key=int(json['key']),
            data=str(json['data']),
        )


@dataclass
class ScanRecord:
    '''
    Stores the byte data of the advertisement packet sent by a Bluetooth device.
    '''
    name: typing.Optional[str] = None

    uuids: typing.Optional[typing.List[str]] = None

    #: Stores the external appearance description of the device.
    appearance: typing.Optional[int] = None

    #: Stores the transmission power of a broadcasting device.
    tx_power: typing.Optional[int] = None

    #: Key is the company identifier and the value is an array of bytes of
    #: manufacturer specific data.
    manufacturer_data: typing.Optional[typing.List[ManufacturerData]] = None

    def to_json(self):
        json = dict()
        if self.name is not None:
            json['name'] = self.name
        if self.uuids is not None:
            json['uuids'] = [i for i in self.uuids]
        if self.appearance is not None:
            json['appearance'] = self.appearance
        if self.tx_power is not None:
            json['txPower'] = self.tx_power
        if self.manufacturer_data is not None:
            json['manufacturerData'] = [i.to_json() for i in self.manufacturer_data]
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']) if 'name' in json else None,
            uuids=[str(i) for i in json['uuids']] if 'uuids' in json else None,
            appearance=int(json['appearance']) if 'appearance' in json else None,
            tx_power=int(json['txPower']) if 'txPower' in json else None,
            manufacturer_data=[ManufacturerData.from_json(i) for i in json['manufacturerData']] if 'manufacturerData' in json else None,
        )


@dataclass
class ScanEntry:
    '''
    Stores the advertisement packet information that is sent by a Bluetooth device.
    '''
    device_address: str

    rssi: int

    scan_record: ScanRecord

    def to_json(self):
        json = dict()
        json['deviceAddress'] = self.device_address
        json['rssi'] = self.rssi
        json['scanRecord'] = self.scan_record.to_json()
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            device_address=str(json['deviceAddress']),
            rssi=int(json['rssi']),
            scan_record=ScanRecord.from_json(json['scanRecord']),
        )


def enable(
        state: CentralState
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enable the BluetoothEmulation domain.

    :param state: State of the simulated central.
    '''
    params: T_JSON_DICT = dict()
    params['state'] = state.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'BluetoothEmulation.enable',
        'params': params,
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disable the BluetoothEmulation domain.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'BluetoothEmulation.disable',
    }
    json = yield cmd_dict


def simulate_preconnected_peripheral(
        address: str,
        name: str,
        manufacturer_data: typing.List[ManufacturerData],
        known_service_uuids: typing.List[str]
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Simulates a peripheral with ``address``, ``name`` and ``knownServiceUuids``
    that has already been connected to the system.

    :param address:
    :param name:
    :param manufacturer_data:
    :param known_service_uuids:
    '''
    params: T_JSON_DICT = dict()
    params['address'] = address
    params['name'] = name
    params['manufacturerData'] = [i.to_json() for i in manufacturer_data]
    params['knownServiceUuids'] = [i for i in known_service_uuids]
    cmd_dict: T_JSON_DICT = {
        'method': 'BluetoothEmulation.simulatePreconnectedPeripheral',
        'params': params,
    }
    json = yield cmd_dict


def simulate_advertisement(
        entry: ScanEntry
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Simulates an advertisement packet described in ``entry`` being received by
    the central.

    :param entry:
    '''
    params: T_JSON_DICT = dict()
    params['entry'] = entry.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'BluetoothEmulation.simulateAdvertisement',
        'params': params,
    }
    json = yield cmd_dict

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\sqlite\provision.py ===
# dialects/sqlite/provision.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors

import os
import re

from ... import exc
from ...engine import url as sa_url
from ...testing.provision import create_db
from ...testing.provision import drop_db
from ...testing.provision import follower_url_from_main
from ...testing.provision import generate_driver_url
from ...testing.provision import log
from ...testing.provision import post_configure_engine
from ...testing.provision import run_reap_dbs
from ...testing.provision import stop_test_class_outside_fixtures
from ...testing.provision import temp_table_keyword_args
from ...testing.provision import upsert


# TODO: I can't get this to build dynamically with pytest-xdist procs
_drivernames = {
    "pysqlite",
    "aiosqlite",
    "pysqlcipher",
    "pysqlite_numeric",
    "pysqlite_dollar",
}


def _format_url(url, driver, ident):
    """given a sqlite url + desired driver + ident, make a canonical
    URL out of it

    """
    url = sa_url.make_url(url)

    if driver is None:
        driver = url.get_driver_name()

    filename = url.database

    needs_enc = driver == "pysqlcipher"
    name_token = None

    if filename and filename != ":memory:":
        assert "test_schema" not in filename
        tokens = re.split(r"[_\.]", filename)

        for token in tokens:
            if token in _drivernames:
                if driver is None:
                    driver = token
                continue
            elif token in ("db", "enc"):
                continue
            elif name_token is None:
                name_token = token.strip("_")

        assert name_token, f"sqlite filename has no name token: {url.database}"

        new_filename = f"{name_token}_{driver}"
        if ident:
            new_filename += f"_{ident}"
        new_filename += ".db"
        if needs_enc:
            new_filename += ".enc"
        url = url.set(database=new_filename)

    if needs_enc:
        url = url.set(password="test")

    url = url.set(drivername="sqlite+%s" % (driver,))

    return url


@generate_driver_url.for_db("sqlite")
def generate_driver_url(url, driver, query_str):
    url = _format_url(url, driver, None)

    try:
        url.get_dialect()
    except exc.NoSuchModuleError:
        return None
    else:
        return url


@follower_url_from_main.for_db("sqlite")
def _sqlite_follower_url_from_main(url, ident):
    return _format_url(url, None, ident)


@post_configure_engine.for_db("sqlite")
def _sqlite_post_configure_engine(url, engine, follower_ident):
    from sqlalchemy import event

    if follower_ident:
        attach_path = f"{follower_ident}_{engine.driver}_test_schema.db"
    else:
        attach_path = f"{engine.driver}_test_schema.db"

    @event.listens_for(engine, "connect")
    def connect(dbapi_connection, connection_record):
        # use file DBs in all cases, memory acts kind of strangely
        # as an attached

        # NOTE!  this has to be done *per connection*.  New sqlite connection,
        # as we get with say, QueuePool, the attaches are gone.
        # so schemes to delete those attached files have to be done at the
        # filesystem level and not rely upon what attachments are in a
        # particular SQLite connection
        dbapi_connection.execute(
            f'ATTACH DATABASE "{attach_path}" AS test_schema'
        )

    @event.listens_for(engine, "engine_disposed")
    def dispose(engine):
        """most databases should be dropped using
        stop_test_class_outside_fixtures

        however a few tests like AttachedDBTest might not get triggered on
        that main hook

        """

        if os.path.exists(attach_path):
            os.remove(attach_path)

        filename = engine.url.database

        if filename and filename != ":memory:" and os.path.exists(filename):
            os.remove(filename)


@create_db.for_db("sqlite")
def _sqlite_create_db(cfg, eng, ident):
    pass


@drop_db.for_db("sqlite")
def _sqlite_drop_db(cfg, eng, ident):
    _drop_dbs_w_ident(eng.url.database, eng.driver, ident)


def _drop_dbs_w_ident(databasename, driver, ident):
    for path in os.listdir("."):
        fname, ext = os.path.split(path)
        if ident in fname and ext in [".db", ".db.enc"]:
            log.info("deleting SQLite database file: %s", path)
            os.remove(path)


@stop_test_class_outside_fixtures.for_db("sqlite")
def stop_test_class_outside_fixtures(config, db, cls):
    db.dispose()


@temp_table_keyword_args.for_db("sqlite")
def _sqlite_temp_table_keyword_args(cfg, eng):
    return {"prefixes": ["TEMPORARY"]}


@run_reap_dbs.for_db("sqlite")
def _reap_sqlite_dbs(url, idents):
    log.info("db reaper connecting to %r", url)
    log.info("identifiers in file: %s", ", ".join(idents))
    url = sa_url.make_url(url)
    for ident in idents:
        for drivername in _drivernames:
            _drop_dbs_w_ident(url.database, drivername, ident)


@upsert.for_db("sqlite")
def _upsert(
    cfg, table, returning, *, set_lambda=None, sort_by_parameter_order=False
):
    from sqlalchemy.dialects.sqlite import insert

    stmt = insert(table)

    if set_lambda:
        stmt = stmt.on_conflict_do_update(set_=set_lambda(stmt.excluded))
    else:
        stmt = stmt.on_conflict_do_nothing()

    stmt = stmt.returning(
        *returning, sort_by_parameter_order=sort_by_parameter_order
    )
    return stmt

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\liealgebras\root_system.py ===
from .cartan_type import CartanType
from sympy.core.basic import Atom

class RootSystem(Atom):
    """Represent the root system of a simple Lie algebra

    Every simple Lie algebra has a unique root system.  To find the root
    system, we first consider the Cartan subalgebra of g, which is the maximal
    abelian subalgebra, and consider the adjoint action of g on this
    subalgebra.  There is a root system associated with this action. Now, a
    root system over a vector space V is a set of finite vectors Phi (called
    roots), which satisfy:

    1.  The roots span V
    2.  The only scalar multiples of x in Phi are x and -x
    3.  For every x in Phi, the set Phi is closed under reflection
        through the hyperplane perpendicular to x.
    4.  If x and y are roots in Phi, then the projection of y onto
        the line through x is a half-integral multiple of x.

    Now, there is a subset of Phi, which we will call Delta, such that:
    1.  Delta is a basis of V
    2.  Each root x in Phi can be written x = sum k_y y for y in Delta

    The elements of Delta are called the simple roots.
    Therefore, we see that the simple roots span the root space of a given
    simple Lie algebra.

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Root_system
    .. [2] Lie Algebras and Representation Theory - Humphreys

    """

    def __new__(cls, cartantype):
        """Create a new RootSystem object

        This method assigns an attribute called cartan_type to each instance of
        a RootSystem object.  When an instance of RootSystem is called, it
        needs an argument, which should be an instance of a simple Lie algebra.
        We then take the CartanType of this argument and set it as the
        cartan_type attribute of the RootSystem instance.

        """
        obj = Atom.__new__(cls)
        obj.cartan_type = CartanType(cartantype)
        return obj

    def simple_roots(self):
        """Generate the simple roots of the Lie algebra

        The rank of the Lie algebra determines the number of simple roots that
        it has.  This method obtains the rank of the Lie algebra, and then uses
        the simple_root method from the Lie algebra classes to generate all the
        simple roots.

        Examples
        ========

        >>> from sympy.liealgebras.root_system import RootSystem
        >>> c = RootSystem("A3")
        >>> roots = c.simple_roots()
        >>> roots
        {1: [1, -1, 0, 0], 2: [0, 1, -1, 0], 3: [0, 0, 1, -1]}

        """
        n = self.cartan_type.rank()
        roots = {i: self.cartan_type.simple_root(i) for i in range(1, n+1)}
        return roots


    def all_roots(self):
        """Generate all the roots of a given root system

        The result is a dictionary where the keys are integer numbers.  It
        generates the roots by getting the dictionary of all positive roots
        from the bases classes, and then taking each root, and multiplying it
        by -1 and adding it to the dictionary.  In this way all the negative
        roots are generated.

        """
        alpha = self.cartan_type.positive_roots()
        keys = list(alpha.keys())
        k = max(keys)
        for val in keys:
            k += 1
            root = alpha[val]
            newroot = [-x for x in root]
            alpha[k] = newroot
        return alpha

    def root_space(self):
        """Return the span of the simple roots

        The root space is the vector space spanned by the simple roots, i.e. it
        is a vector space with a distinguished basis, the simple roots.  This
        method returns a string that represents the root space as the span of
        the simple roots, alpha[1],...., alpha[n].

        Examples
        ========

        >>> from sympy.liealgebras.root_system import RootSystem
        >>> c = RootSystem("A3")
        >>> c.root_space()
        'alpha[1] + alpha[2] + alpha[3]'

        """
        n = self.cartan_type.rank()
        rs = " + ".join("alpha["+str(i) +"]" for i in range(1, n+1))
        return rs

    def add_simple_roots(self, root1, root2):
        """Add two simple roots together

        The function takes as input two integers, root1 and root2.  It then
        uses these integers as keys in the dictionary of simple roots, and gets
        the corresponding simple roots, and then adds them together.

        Examples
        ========

        >>> from sympy.liealgebras.root_system import RootSystem
        >>> c = RootSystem("A3")
        >>> newroot = c.add_simple_roots(1, 2)
        >>> newroot
        [1, 0, -1, 0]

        """

        alpha = self.simple_roots()
        if root1 > len(alpha) or root2 > len(alpha):
            raise ValueError("You've used a root that doesn't exist!")
        a1 = alpha[root1]
        a2 = alpha[root2]
        newroot = [_a1 + _a2 for _a1, _a2 in zip(a1, a2)]
        return newroot

    def add_as_roots(self, root1, root2):
        """Add two roots together if and only if their sum is also a root

        It takes as input two vectors which should be roots.  It then computes
        their sum and checks if it is in the list of all possible roots.  If it
        is, it returns the sum.  Otherwise it returns a string saying that the
        sum is not a root.

        Examples
        ========

        >>> from sympy.liealgebras.root_system import RootSystem
        >>> c = RootSystem("A3")
        >>> c.add_as_roots([1, 0, -1, 0], [0, 0, 1, -1])
        [1, 0, 0, -1]
        >>> c.add_as_roots([1, -1, 0, 0], [0, 0, -1, 1])
        'The sum of these two roots is not a root'

        """
        alpha = self.all_roots()
        newroot = [r1 + r2 for r1, r2 in zip(root1, root2)]
        if newroot in alpha.values():
            return newroot
        else:
            return "The sum of these two roots is not a root"


    def cartan_matrix(self):
        """Cartan matrix of Lie algebra associated with this root system

        Examples
        ========

        >>> from sympy.liealgebras.root_system import RootSystem
        >>> c = RootSystem("A3")
        >>> c.cartan_matrix()
        Matrix([
            [ 2, -1,  0],
            [-1,  2, -1],
            [ 0, -1,  2]])
        """
        return self.cartan_type.cartan_matrix()

    def dynkin_diagram(self):
        """Dynkin diagram of the Lie algebra associated with this root system

        Examples
        ========

        >>> from sympy.liealgebras.root_system import RootSystem
        >>> c = RootSystem("A3")
        >>> print(c.dynkin_diagram())
        0---0---0
        1   2   3
        """
        return self.cartan_type.dynkin_diagram()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\matrices\immutable.py ===
from mpmath.matrices.matrices import _matrix

from sympy.core import Basic, Dict, Tuple
from sympy.core.numbers import Integer
from sympy.core.cache import cacheit
from sympy.core.sympify import _sympy_converter as sympify_converter, _sympify
from sympy.matrices.dense import DenseMatrix
from sympy.matrices.expressions import MatrixExpr
from sympy.matrices.matrixbase import MatrixBase
from sympy.matrices.repmatrix import RepMatrix
from sympy.matrices.sparse import SparseRepMatrix
from sympy.multipledispatch import dispatch


def sympify_matrix(arg):
    return arg.as_immutable()


sympify_converter[MatrixBase] = sympify_matrix


def sympify_mpmath_matrix(arg):
    mat = [_sympify(x) for x in arg]
    return ImmutableDenseMatrix(arg.rows, arg.cols, mat)


sympify_converter[_matrix] = sympify_mpmath_matrix


class ImmutableRepMatrix(RepMatrix, MatrixExpr): # type: ignore
    """Immutable matrix based on RepMatrix

    Uses DomainMAtrix as the internal representation.
    """

    #
    # This is a subclass of RepMatrix that adds/overrides some methods to make
    # the instances Basic and immutable. ImmutableRepMatrix is a superclass for
    # both ImmutableDenseMatrix and ImmutableSparseMatrix.
    #

    def __new__(cls, *args, **kwargs):
        return cls._new(*args, **kwargs)

    __hash__ = MatrixExpr.__hash__

    def copy(self):
        return self

    @property
    def cols(self):
        return self._cols

    @property
    def rows(self):
        return self._rows

    @property
    def shape(self):
        return self._rows, self._cols

    def as_immutable(self):
        return self

    def _entry(self, i, j, **kwargs):
        return self[i, j]

    def __setitem__(self, *args):
        raise TypeError("Cannot set values of {}".format(self.__class__))

    def is_diagonalizable(self, reals_only=False, **kwargs):
        return super().is_diagonalizable(
            reals_only=reals_only, **kwargs)

    is_diagonalizable.__doc__ = SparseRepMatrix.is_diagonalizable.__doc__
    is_diagonalizable = cacheit(is_diagonalizable)

    def analytic_func(self, f, x):
        return self.as_mutable().analytic_func(f, x).as_immutable()


class ImmutableDenseMatrix(DenseMatrix, ImmutableRepMatrix):  # type: ignore
    """Create an immutable version of a matrix.

    Examples
    ========

    >>> from sympy import eye, ImmutableMatrix
    >>> ImmutableMatrix(eye(3))
    Matrix([
    [1, 0, 0],
    [0, 1, 0],
    [0, 0, 1]])
    >>> _[0, 0] = 42
    Traceback (most recent call last):
    ...
    TypeError: Cannot set values of ImmutableDenseMatrix
    """

    # MatrixExpr is set as NotIterable, but we want explicit matrices to be
    # iterable
    _iterable = True
    _class_priority = 8
    _op_priority = 10.001

    @classmethod
    def _new(cls, *args, **kwargs):
        if len(args) == 1 and isinstance(args[0], ImmutableDenseMatrix):
            return args[0]
        if kwargs.get('copy', True) is False:
            if len(args) != 3:
                raise TypeError("'copy=False' requires a matrix be initialized as rows,cols,[list]")
            rows, cols, flat_list = args
        else:
            rows, cols, flat_list = cls._handle_creation_inputs(*args, **kwargs)
            flat_list = list(flat_list) # create a shallow copy

        rep = cls._flat_list_to_DomainMatrix(rows, cols, flat_list)

        return cls._fromrep(rep)

    @classmethod
    def _fromrep(cls, rep):
        rows, cols = rep.shape
        flat_list = rep.to_sympy().to_list_flat()
        obj = Basic.__new__(cls,
            Integer(rows),
            Integer(cols),
            Tuple(*flat_list, sympify=False))
        obj._rows = rows
        obj._cols = cols
        obj._rep = rep
        return obj


# make sure ImmutableDenseMatrix is aliased as ImmutableMatrix
ImmutableMatrix = ImmutableDenseMatrix


class ImmutableSparseMatrix(SparseRepMatrix, ImmutableRepMatrix):  # type:ignore
    """Create an immutable version of a sparse matrix.

    Examples
    ========

    >>> from sympy import eye, ImmutableSparseMatrix
    >>> ImmutableSparseMatrix(1, 1, {})
    Matrix([[0]])
    >>> ImmutableSparseMatrix(eye(3))
    Matrix([
    [1, 0, 0],
    [0, 1, 0],
    [0, 0, 1]])
    >>> _[0, 0] = 42
    Traceback (most recent call last):
    ...
    TypeError: Cannot set values of ImmutableSparseMatrix
    >>> _.shape
    (3, 3)
    """
    is_Matrix = True
    _class_priority = 9

    @classmethod
    def _new(cls, *args, **kwargs):
        rows, cols, smat = cls._handle_creation_inputs(*args, **kwargs)

        rep = cls._smat_to_DomainMatrix(rows, cols, smat)

        return cls._fromrep(rep)

    @classmethod
    def _fromrep(cls, rep):
        rows, cols = rep.shape
        smat = rep.to_sympy().to_dok()
        obj = Basic.__new__(cls, Integer(rows), Integer(cols), Dict(smat))
        obj._rows = rows
        obj._cols = cols
        obj._rep = rep
        return obj


@dispatch(ImmutableDenseMatrix, ImmutableDenseMatrix)
def _eval_is_eq(lhs, rhs): # noqa:F811
    """Helper method for Equality with matrices.sympy.

    Relational automatically converts matrices to ImmutableDenseMatrix
    instances, so this method only applies here.  Returns True if the
    matrices are definitively the same, False if they are definitively
    different, and None if undetermined (e.g. if they contain Symbols).
    Returning None triggers default handling of Equalities.

    """
    if lhs.shape != rhs.shape:
        return False
    return (lhs - rhs).is_zero_matrix

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\tensor\array\sparse_ndim_array.py ===
from sympy.core.basic import Basic
from sympy.core.containers import (Dict, Tuple)
from sympy.core.singleton import S
from sympy.core.sympify import _sympify
from sympy.tensor.array.mutable_ndim_array import MutableNDimArray
from sympy.tensor.array.ndim_array import NDimArray, ImmutableNDimArray
from sympy.utilities.iterables import flatten

import functools

class SparseNDimArray(NDimArray):

    def __new__(self, *args, **kwargs):
        return ImmutableSparseNDimArray(*args, **kwargs)

    def __getitem__(self, index):
        """
        Get an element from a sparse N-dim array.

        Examples
        ========

        >>> from sympy import MutableSparseNDimArray
        >>> a = MutableSparseNDimArray(range(4), (2, 2))
        >>> a
        [[0, 1], [2, 3]]
        >>> a[0, 0]
        0
        >>> a[1, 1]
        3
        >>> a[0]
        [0, 1]
        >>> a[1]
        [2, 3]

        Symbolic indexing:

        >>> from sympy.abc import i, j
        >>> a[i, j]
        [[0, 1], [2, 3]][i, j]

        Replace `i` and `j` to get element `(0, 0)`:

        >>> a[i, j].subs({i: 0, j: 0})
        0

        """
        syindex = self._check_symbolic_index(index)
        if syindex is not None:
            return syindex

        index = self._check_index_for_getitem(index)

        # `index` is a tuple with one or more slices:
        if isinstance(index, tuple) and any(isinstance(i, slice) for i in index):
            sl_factors, eindices = self._get_slice_data_for_array_access(index)
            array = [self._sparse_array.get(self._parse_index(i), S.Zero) for i in eindices]
            nshape = [len(el) for i, el in enumerate(sl_factors) if isinstance(index[i], slice)]
            return type(self)(array, nshape)
        else:
            index = self._parse_index(index)
            return self._sparse_array.get(index, S.Zero)

    @classmethod
    def zeros(cls, *shape):
        """
        Return a sparse N-dim array of zeros.
        """
        return cls({}, shape)

    def tomatrix(self):
        """
        Converts MutableDenseNDimArray to Matrix. Can convert only 2-dim array, else will raise error.

        Examples
        ========

        >>> from sympy import MutableSparseNDimArray
        >>> a = MutableSparseNDimArray([1 for i in range(9)], (3, 3))
        >>> b = a.tomatrix()
        >>> b
        Matrix([
        [1, 1, 1],
        [1, 1, 1],
        [1, 1, 1]])
        """
        from sympy.matrices import SparseMatrix
        if self.rank() != 2:
            raise ValueError('Dimensions must be of size of 2')

        mat_sparse = {}
        for key, value in self._sparse_array.items():
            mat_sparse[self._get_tuple_index(key)] = value

        return SparseMatrix(self.shape[0], self.shape[1], mat_sparse)

    def reshape(self, *newshape):
        new_total_size = functools.reduce(lambda x,y: x*y, newshape)
        if new_total_size != self._loop_size:
            raise ValueError("Invalid reshape parameters " + newshape)

        return type(self)(self._sparse_array, newshape)

class ImmutableSparseNDimArray(SparseNDimArray, ImmutableNDimArray): # type: ignore

    def __new__(cls, iterable=None, shape=None, **kwargs):
        shape, flat_list = cls._handle_ndarray_creation_inputs(iterable, shape, **kwargs)
        shape = Tuple(*map(_sympify, shape))
        cls._check_special_bounds(flat_list, shape)
        loop_size = functools.reduce(lambda x,y: x*y, shape) if shape else len(flat_list)

        # Sparse array:
        if isinstance(flat_list, (dict, Dict)):
            sparse_array = Dict(flat_list)
        else:
            sparse_array = {}
            for i, el in enumerate(flatten(flat_list)):
                if el != 0:
                    sparse_array[i] = _sympify(el)

        sparse_array = Dict(sparse_array)

        self = Basic.__new__(cls, sparse_array, shape, **kwargs)
        self._shape = shape
        self._rank = len(shape)
        self._loop_size = loop_size
        self._sparse_array = sparse_array

        return self

    def __setitem__(self, index, value):
        raise TypeError("immutable N-dim array")

    def as_mutable(self):
        return MutableSparseNDimArray(self)


class MutableSparseNDimArray(MutableNDimArray, SparseNDimArray):

    def __new__(cls, iterable=None, shape=None, **kwargs):
        shape, flat_list = cls._handle_ndarray_creation_inputs(iterable, shape, **kwargs)
        self = object.__new__(cls)
        self._shape = shape
        self._rank = len(shape)
        self._loop_size = functools.reduce(lambda x,y: x*y, shape) if shape else len(flat_list)

        # Sparse array:
        if isinstance(flat_list, (dict, Dict)):
            self._sparse_array = dict(flat_list)
            return self

        self._sparse_array = {}

        for i, el in enumerate(flatten(flat_list)):
            if el != 0:
                self._sparse_array[i] = _sympify(el)

        return self

    def __setitem__(self, index, value):
        """Allows to set items to MutableDenseNDimArray.

        Examples
        ========

        >>> from sympy import MutableSparseNDimArray
        >>> a = MutableSparseNDimArray.zeros(2, 2)
        >>> a[0, 0] = 1
        >>> a[1, 1] = 1
        >>> a
        [[1, 0], [0, 1]]
        """
        if isinstance(index, tuple) and any(isinstance(i, slice) for i in index):
            value, eindices, slice_offsets = self._get_slice_data_for_array_assignment(index, value)
            for i in eindices:
                other_i = [ind - j for ind, j in zip(i, slice_offsets) if j is not None]
                other_value = value[other_i]
                complete_index = self._parse_index(i)
                if other_value != 0:
                    self._sparse_array[complete_index] = other_value
                elif complete_index in self._sparse_array:
                    self._sparse_array.pop(complete_index)
        else:
            index = self._parse_index(index)
            value = _sympify(value)
            if value == 0 and index in self._sparse_array:
                self._sparse_array.pop(index)
            else:
                self._sparse_array[index] = value

    def as_immutable(self):
        return ImmutableSparseNDimArray(self)

    @property
    def free_symbols(self):
        return {i for j in self._sparse_array.values() for i in j.free_symbols}

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\colorama\winterm.py ===
# Copyright Jonathan Hartley 2013. BSD 3-Clause license, see LICENSE file.
try:
    from msvcrt import get_osfhandle
except ImportError:
    def get_osfhandle(_):
        raise OSError("This isn't windows!")


from . import win32

# from wincon.h
class WinColor(object):
    BLACK   = 0
    BLUE    = 1
    GREEN   = 2
    CYAN    = 3
    RED     = 4
    MAGENTA = 5
    YELLOW  = 6
    GREY    = 7

# from wincon.h
class WinStyle(object):
    NORMAL              = 0x00 # dim text, dim background
    BRIGHT              = 0x08 # bright text, dim background
    BRIGHT_BACKGROUND   = 0x80 # dim text, bright background

class WinTerm(object):

    def __init__(self):
        self._default = win32.GetConsoleScreenBufferInfo(win32.STDOUT).wAttributes
        self.set_attrs(self._default)
        self._default_fore = self._fore
        self._default_back = self._back
        self._default_style = self._style
        # In order to emulate LIGHT_EX in windows, we borrow the BRIGHT style.
        # So that LIGHT_EX colors and BRIGHT style do not clobber each other,
        # we track them separately, since LIGHT_EX is overwritten by Fore/Back
        # and BRIGHT is overwritten by Style codes.
        self._light = 0

    def get_attrs(self):
        return self._fore + self._back * 16 + (self._style | self._light)

    def set_attrs(self, value):
        self._fore = value & 7
        self._back = (value >> 4) & 7
        self._style = value & (WinStyle.BRIGHT | WinStyle.BRIGHT_BACKGROUND)

    def reset_all(self, on_stderr=None):
        self.set_attrs(self._default)
        self.set_console(attrs=self._default)
        self._light = 0

    def fore(self, fore=None, light=False, on_stderr=False):
        if fore is None:
            fore = self._default_fore
        self._fore = fore
        # Emulate LIGHT_EX with BRIGHT Style
        if light:
            self._light |= WinStyle.BRIGHT
        else:
            self._light &= ~WinStyle.BRIGHT
        self.set_console(on_stderr=on_stderr)

    def back(self, back=None, light=False, on_stderr=False):
        if back is None:
            back = self._default_back
        self._back = back
        # Emulate LIGHT_EX with BRIGHT_BACKGROUND Style
        if light:
            self._light |= WinStyle.BRIGHT_BACKGROUND
        else:
            self._light &= ~WinStyle.BRIGHT_BACKGROUND
        self.set_console(on_stderr=on_stderr)

    def style(self, style=None, on_stderr=False):
        if style is None:
            style = self._default_style
        self._style = style
        self.set_console(on_stderr=on_stderr)

    def set_console(self, attrs=None, on_stderr=False):
        if attrs is None:
            attrs = self.get_attrs()
        handle = win32.STDOUT
        if on_stderr:
            handle = win32.STDERR
        win32.SetConsoleTextAttribute(handle, attrs)

    def get_position(self, handle):
        position = win32.GetConsoleScreenBufferInfo(handle).dwCursorPosition
        # Because Windows coordinates are 0-based,
        # and win32.SetConsoleCursorPosition expects 1-based.
        position.X += 1
        position.Y += 1
        return position

    def set_cursor_position(self, position=None, on_stderr=False):
        if position is None:
            # I'm not currently tracking the position, so there is no default.
            # position = self.get_position()
            return
        handle = win32.STDOUT
        if on_stderr:
            handle = win32.STDERR
        win32.SetConsoleCursorPosition(handle, position)

    def cursor_adjust(self, x, y, on_stderr=False):
        handle = win32.STDOUT
        if on_stderr:
            handle = win32.STDERR
        position = self.get_position(handle)
        adjusted_position = (position.Y + y, position.X + x)
        win32.SetConsoleCursorPosition(handle, adjusted_position, adjust=False)

    def erase_screen(self, mode=0, on_stderr=False):
        # 0 should clear from the cursor to the end of the screen.
        # 1 should clear from the cursor to the beginning of the screen.
        # 2 should clear the entire screen, and move cursor to (1,1)
        handle = win32.STDOUT
        if on_stderr:
            handle = win32.STDERR
        csbi = win32.GetConsoleScreenBufferInfo(handle)
        # get the number of character cells in the current buffer
        cells_in_screen = csbi.dwSize.X * csbi.dwSize.Y
        # get number of character cells before current cursor position
        cells_before_cursor = csbi.dwSize.X * csbi.dwCursorPosition.Y + csbi.dwCursorPosition.X
        if mode == 0:
            from_coord = csbi.dwCursorPosition
            cells_to_erase = cells_in_screen - cells_before_cursor
        elif mode == 1:
            from_coord = win32.COORD(0, 0)
            cells_to_erase = cells_before_cursor
        elif mode == 2:
            from_coord = win32.COORD(0, 0)
            cells_to_erase = cells_in_screen
        else:
            # invalid mode
            return
        # fill the entire screen with blanks
        win32.FillConsoleOutputCharacter(handle, ' ', cells_to_erase, from_coord)
        # now set the buffer's attributes accordingly
        win32.FillConsoleOutputAttribute(handle, self.get_attrs(), cells_to_erase, from_coord)
        if mode == 2:
            # put the cursor where needed
            win32.SetConsoleCursorPosition(handle, (1, 1))

    def erase_line(self, mode=0, on_stderr=False):
        # 0 should clear from the cursor to the end of the line.
        # 1 should clear from the cursor to the beginning of the line.
        # 2 should clear the entire line.
        handle = win32.STDOUT
        if on_stderr:
            handle = win32.STDERR
        csbi = win32.GetConsoleScreenBufferInfo(handle)
        if mode == 0:
            from_coord = csbi.dwCursorPosition
            cells_to_erase = csbi.dwSize.X - csbi.dwCursorPosition.X
        elif mode == 1:
            from_coord = win32.COORD(0, csbi.dwCursorPosition.Y)
            cells_to_erase = csbi.dwCursorPosition.X
        elif mode == 2:
            from_coord = win32.COORD(0, csbi.dwCursorPosition.Y)
            cells_to_erase = csbi.dwSize.X
        else:
            # invalid mode
            return
        # fill the entire screen with blanks
        win32.FillConsoleOutputCharacter(handle, ' ', cells_to_erase, from_coord)
        # now set the buffer's attributes accordingly
        win32.FillConsoleOutputAttribute(handle, self.get_attrs(), cells_to_erase, from_coord)

    def set_title(self, title):
        win32.SetConsoleTitle(title)


def enable_vt_processing(fd):
    if win32.windll is None or not win32.winapi_test():
        return False

    try:
        handle = get_osfhandle(fd)
        mode = win32.GetConsoleMode(handle)
        win32.SetConsoleMode(
            handle,
            mode | win32.ENABLE_VIRTUAL_TERMINAL_PROCESSING,
        )

        mode = win32.GetConsoleMode(handle)
        if mode & win32.ENABLE_VIRTUAL_TERMINAL_PROCESSING:
            return True
    # Can get TypeError in testsuite where 'fd' is a Mock()
    except (OSError, TypeError):
        return False

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\typing\mypy_plugin.py ===
"""A mypy_ plugin for managing a number of platform-specific annotations.
Its functionality can be split into three distinct parts:

* Assigning the (platform-dependent) precisions of certain `~numpy.number`
  subclasses, including the likes of `~numpy.int_`, `~numpy.intp` and
  `~numpy.longlong`. See the documentation on
  :ref:`scalar types <arrays.scalars.built-in>` for a comprehensive overview
  of the affected classes. Without the plugin the precision of all relevant
  classes will be inferred as `~typing.Any`.
* Removing all extended-precision `~numpy.number` subclasses that are
  unavailable for the platform in question. Most notably this includes the
  likes of `~numpy.float128` and `~numpy.complex256`. Without the plugin *all*
  extended-precision types will, as far as mypy is concerned, be available
  to all platforms.
* Assigning the (platform-dependent) precision of `~numpy.ctypeslib.c_intp`.
  Without the plugin the type will default to `ctypes.c_int64`.

  .. versionadded:: 1.22

.. deprecated:: 2.3

Examples
--------
To enable the plugin, one must add it to their mypy `configuration file`_:

.. code-block:: ini

    [mypy]
    plugins = numpy.typing.mypy_plugin

.. _mypy: https://mypy-lang.org/
.. _configuration file: https://mypy.readthedocs.io/en/stable/config_file.html

"""

from collections.abc import Callable, Iterable
from typing import TYPE_CHECKING, Final, TypeAlias, cast

import numpy as np

__all__: list[str] = []


def _get_precision_dict() -> dict[str, str]:
    names = [
        ("_NBitByte", np.byte),
        ("_NBitShort", np.short),
        ("_NBitIntC", np.intc),
        ("_NBitIntP", np.intp),
        ("_NBitInt", np.int_),
        ("_NBitLong", np.long),
        ("_NBitLongLong", np.longlong),

        ("_NBitHalf", np.half),
        ("_NBitSingle", np.single),
        ("_NBitDouble", np.double),
        ("_NBitLongDouble", np.longdouble),
    ]
    ret: dict[str, str] = {}
    for name, typ in names:
        n = 8 * np.dtype(typ).itemsize
        ret[f"{_MODULE}._nbit.{name}"] = f"{_MODULE}._nbit_base._{n}Bit"
    return ret


def _get_extended_precision_list() -> list[str]:
    extended_names = [
        "float96",
        "float128",
        "complex192",
        "complex256",
    ]
    return [i for i in extended_names if hasattr(np, i)]

def _get_c_intp_name() -> str:
    # Adapted from `np.core._internal._getintp_ctype`
    return {
        "i": "c_int",
        "l": "c_long",
        "q": "c_longlong",
    }.get(np.dtype("n").char, "c_long")


_MODULE: Final = "numpy._typing"

#: A dictionary mapping type-aliases in `numpy._typing._nbit` to
#: concrete `numpy.typing.NBitBase` subclasses.
_PRECISION_DICT: Final = _get_precision_dict()

#: A list with the names of all extended precision `np.number` subclasses.
_EXTENDED_PRECISION_LIST: Final = _get_extended_precision_list()

#: The name of the ctypes equivalent of `np.intp`
_C_INTP: Final = _get_c_intp_name()


try:
    if TYPE_CHECKING:
        from mypy.typeanal import TypeAnalyser

    import mypy.types
    from mypy.build import PRI_MED
    from mypy.nodes import ImportFrom, MypyFile, Statement
    from mypy.plugin import AnalyzeTypeContext, Plugin

except ModuleNotFoundError as e:

    def plugin(version: str) -> type:
        raise e

else:

    _HookFunc: TypeAlias = Callable[[AnalyzeTypeContext], mypy.types.Type]

    def _hook(ctx: AnalyzeTypeContext) -> mypy.types.Type:
        """Replace a type-alias with a concrete ``NBitBase`` subclass."""
        typ, _, api = ctx
        name = typ.name.split(".")[-1]
        name_new = _PRECISION_DICT[f"{_MODULE}._nbit.{name}"]
        return cast("TypeAnalyser", api).named_type(name_new)

    def _index(iterable: Iterable[Statement], id: str) -> int:
        """Identify the first ``ImportFrom`` instance the specified `id`."""
        for i, value in enumerate(iterable):
            if getattr(value, "id", None) == id:
                return i
        raise ValueError("Failed to identify a `ImportFrom` instance "
                         f"with the following id: {id!r}")

    def _override_imports(
        file: MypyFile,
        module: str,
        imports: list[tuple[str, str | None]],
    ) -> None:
        """Override the first `module`-based import with new `imports`."""
        # Construct a new `from module import y` statement
        import_obj = ImportFrom(module, 0, names=imports)
        import_obj.is_top_level = True

        # Replace the first `module`-based import statement with `import_obj`
        for lst in [file.defs, cast("list[Statement]", file.imports)]:
            i = _index(lst, module)
            lst[i] = import_obj

    class _NumpyPlugin(Plugin):
        """A mypy plugin for handling versus numpy-specific typing tasks."""

        def get_type_analyze_hook(self, fullname: str) -> _HookFunc | None:
            """Set the precision of platform-specific `numpy.number`
            subclasses.

            For example: `numpy.int_`, `numpy.longlong` and `numpy.longdouble`.
            """
            if fullname in _PRECISION_DICT:
                return _hook
            return None

        def get_additional_deps(
            self, file: MypyFile
        ) -> list[tuple[int, str, int]]:
            """Handle all import-based overrides.

            * Import platform-specific extended-precision `numpy.number`
              subclasses (*e.g.* `numpy.float96` and `numpy.float128`).
            * Import the appropriate `ctypes` equivalent to `numpy.intp`.

            """
            fullname = file.fullname
            if fullname == "numpy":
                _override_imports(
                    file,
                    f"{_MODULE}._extended_precision",
                    imports=[(v, v) for v in _EXTENDED_PRECISION_LIST],
                )
            elif fullname == "numpy.ctypeslib":
                _override_imports(
                    file,
                    "ctypes",
                    imports=[(_C_INTP, "_c_intp")],
                )
            return [(PRI_MED, fullname, -1)]

    def plugin(version: str) -> type:
        import warnings

        plugin = "numpy.typing.mypy_plugin"
        # Deprecated 2025-01-10, NumPy 2.3
        warn_msg = (
            f"`{plugin}` is deprecated, and will be removed in a future "
            f"release. Please remove `plugins = {plugin}` in your mypy config."
            f"(deprecated in NumPy 2.3)"
        )
        warnings.warn(warn_msg, DeprecationWarning, stacklevel=3)

        return _NumpyPlugin

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\chart\chartspace.py ===

# Copyright (c) 2010-2024 openpyxl

"""
Enclosing chart object. The various chart types are actually child objects.
Will probably need to call this indirectly
"""

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Typed,
    String,
    Alias,
)
from openpyxl.descriptors.excel import (
    ExtensionList,
    Relation
)
from openpyxl.descriptors.nested import (
    NestedBool,
    NestedNoneSet,
    NestedString,
    NestedMinMax,
)
from openpyxl.descriptors.sequence import NestedSequence
from openpyxl.xml.constants import CHART_NS

from openpyxl.drawing.colors import ColorMapping
from .text import RichText
from .shapes import GraphicalProperties
from .legend import Legend
from ._3d import _3DBase
from .plotarea import PlotArea
from .title import Title
from .pivot import (
    PivotFormat,
    PivotSource,
)
from .print_settings import PrintSettings


class ChartContainer(Serialisable):

    tagname = "chart"

    title = Typed(expected_type=Title, allow_none=True)
    autoTitleDeleted = NestedBool(allow_none=True)
    pivotFmts = NestedSequence(expected_type=PivotFormat)
    view3D = _3DBase.view3D
    floor = _3DBase.floor
    sideWall = _3DBase.sideWall
    backWall = _3DBase.backWall
    plotArea = Typed(expected_type=PlotArea, )
    legend = Typed(expected_type=Legend, allow_none=True)
    plotVisOnly = NestedBool()
    dispBlanksAs = NestedNoneSet(values=(['span', 'gap', 'zero']))
    showDLblsOverMax = NestedBool(allow_none=True)
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('title', 'autoTitleDeleted', 'pivotFmts', 'view3D',
                    'floor', 'sideWall', 'backWall', 'plotArea', 'legend', 'plotVisOnly',
                    'dispBlanksAs', 'showDLblsOverMax')

    def __init__(self,
                 title=None,
                 autoTitleDeleted=None,
                 pivotFmts=(),
                 view3D=None,
                 floor=None,
                 sideWall=None,
                 backWall=None,
                 plotArea=None,
                 legend=None,
                 plotVisOnly=True,
                 dispBlanksAs="gap",
                 showDLblsOverMax=None,
                 extLst=None,
                ):
        self.title = title
        self.autoTitleDeleted = autoTitleDeleted
        self.pivotFmts = pivotFmts
        self.view3D = view3D
        self.floor = floor
        self.sideWall = sideWall
        self.backWall = backWall
        if plotArea is None:
            plotArea = PlotArea()
        self.plotArea = plotArea
        self.legend = legend
        self.plotVisOnly = plotVisOnly
        self.dispBlanksAs = dispBlanksAs
        self.showDLblsOverMax = showDLblsOverMax


class Protection(Serialisable):

    tagname = "protection"

    chartObject = NestedBool(allow_none=True)
    data = NestedBool(allow_none=True)
    formatting = NestedBool(allow_none=True)
    selection = NestedBool(allow_none=True)
    userInterface = NestedBool(allow_none=True)

    __elements__ = ("chartObject", "data", "formatting", "selection", "userInterface")

    def __init__(self,
                 chartObject=None,
                 data=None,
                 formatting=None,
                 selection=None,
                 userInterface=None,
                ):
        self.chartObject = chartObject
        self.data = data
        self.formatting = formatting
        self.selection = selection
        self.userInterface = userInterface


class ExternalData(Serialisable):

    tagname = "externalData"

    autoUpdate = NestedBool(allow_none=True)
    id = String() # Needs namespace

    def __init__(self,
                 autoUpdate=None,
                 id=None
                ):
        self.autoUpdate = autoUpdate
        self.id = id


class ChartSpace(Serialisable):

    tagname = "chartSpace"

    date1904 = NestedBool(allow_none=True)
    lang = NestedString(allow_none=True)
    roundedCorners = NestedBool(allow_none=True)
    style = NestedMinMax(allow_none=True, min=1, max=48)
    clrMapOvr = Typed(expected_type=ColorMapping, allow_none=True)
    pivotSource = Typed(expected_type=PivotSource, allow_none=True)
    protection = Typed(expected_type=Protection, allow_none=True)
    chart = Typed(expected_type=ChartContainer)
    spPr = Typed(expected_type=GraphicalProperties, allow_none=True)
    graphical_properties = Alias("spPr")
    txPr = Typed(expected_type=RichText, allow_none=True)
    textProperties = Alias("txPr")
    externalData = Typed(expected_type=ExternalData, allow_none=True)
    printSettings = Typed(expected_type=PrintSettings, allow_none=True)
    userShapes = Relation()
    extLst = Typed(expected_type=ExtensionList, allow_none=True)

    __elements__ = ('date1904', 'lang', 'roundedCorners', 'style',
                    'clrMapOvr', 'pivotSource', 'protection', 'chart', 'spPr', 'txPr',
                    'externalData', 'printSettings', 'userShapes')

    def __init__(self,
                 date1904=None,
                 lang=None,
                 roundedCorners=None,
                 style=None,
                 clrMapOvr=None,
                 pivotSource=None,
                 protection=None,
                 chart=None,
                 spPr=None,
                 txPr=None,
                 externalData=None,
                 printSettings=None,
                 userShapes=None,
                 extLst=None,
                ):
        self.date1904 = date1904
        self.lang = lang
        self.roundedCorners = roundedCorners
        self.style = style
        self.clrMapOvr = clrMapOvr
        self.pivotSource = pivotSource
        self.protection = protection
        self.chart = chart
        self.spPr = spPr
        self.txPr = txPr
        self.externalData = externalData
        self.printSettings = printSettings
        self.userShapes = userShapes


    def to_tree(self, tagname=None, idx=None, namespace=None):
        tree = super().to_tree()
        tree.set("xmlns", CHART_NS)
        return tree

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\packaging\_tokenizer.py ===
from __future__ import annotations

import contextlib
import re
from dataclasses import dataclass
from typing import Iterator, NoReturn

from .specifiers import Specifier


@dataclass
class Token:
    name: str
    text: str
    position: int


class ParserSyntaxError(Exception):
    """The provided source text could not be parsed correctly."""

    def __init__(
        self,
        message: str,
        *,
        source: str,
        span: tuple[int, int],
    ) -> None:
        self.span = span
        self.message = message
        self.source = source

        super().__init__()

    def __str__(self) -> str:
        marker = " " * self.span[0] + "~" * (self.span[1] - self.span[0]) + "^"
        return "\n    ".join([self.message, self.source, marker])


DEFAULT_RULES: dict[str, str | re.Pattern[str]] = {
    "LEFT_PARENTHESIS": r"\(",
    "RIGHT_PARENTHESIS": r"\)",
    "LEFT_BRACKET": r"\[",
    "RIGHT_BRACKET": r"\]",
    "SEMICOLON": r";",
    "COMMA": r",",
    "QUOTED_STRING": re.compile(
        r"""
            (
                ('[^']*')
                |
                ("[^"]*")
            )
        """,
        re.VERBOSE,
    ),
    "OP": r"(===|==|~=|!=|<=|>=|<|>)",
    "BOOLOP": r"\b(or|and)\b",
    "IN": r"\bin\b",
    "NOT": r"\bnot\b",
    "VARIABLE": re.compile(
        r"""
            \b(
                python_version
                |python_full_version
                |os[._]name
                |sys[._]platform
                |platform_(release|system)
                |platform[._](version|machine|python_implementation)
                |python_implementation
                |implementation_(name|version)
                |extras?
                |dependency_groups
            )\b
        """,
        re.VERBOSE,
    ),
    "SPECIFIER": re.compile(
        Specifier._operator_regex_str + Specifier._version_regex_str,
        re.VERBOSE | re.IGNORECASE,
    ),
    "AT": r"\@",
    "URL": r"[^ \t]+",
    "IDENTIFIER": r"\b[a-zA-Z0-9][a-zA-Z0-9._-]*\b",
    "VERSION_PREFIX_TRAIL": r"\.\*",
    "VERSION_LOCAL_LABEL_TRAIL": r"\+[a-z0-9]+(?:[-_\.][a-z0-9]+)*",
    "WS": r"[ \t]+",
    "END": r"$",
}


class Tokenizer:
    """Context-sensitive token parsing.

    Provides methods to examine the input stream to check whether the next token
    matches.
    """

    def __init__(
        self,
        source: str,
        *,
        rules: dict[str, str | re.Pattern[str]],
    ) -> None:
        self.source = source
        self.rules: dict[str, re.Pattern[str]] = {
            name: re.compile(pattern) for name, pattern in rules.items()
        }
        self.next_token: Token | None = None
        self.position = 0

    def consume(self, name: str) -> None:
        """Move beyond provided token name, if at current position."""
        if self.check(name):
            self.read()

    def check(self, name: str, *, peek: bool = False) -> bool:
        """Check whether the next token has the provided name.

        By default, if the check succeeds, the token *must* be read before
        another check. If `peek` is set to `True`, the token is not loaded and
        would need to be checked again.
        """
        assert self.next_token is None, (
            f"Cannot check for {name!r}, already have {self.next_token!r}"
        )
        assert name in self.rules, f"Unknown token name: {name!r}"

        expression = self.rules[name]

        match = expression.match(self.source, self.position)
        if match is None:
            return False
        if not peek:
            self.next_token = Token(name, match[0], self.position)
        return True

    def expect(self, name: str, *, expected: str) -> Token:
        """Expect a certain token name next, failing with a syntax error otherwise.

        The token is *not* read.
        """
        if not self.check(name):
            raise self.raise_syntax_error(f"Expected {expected}")
        return self.read()

    def read(self) -> Token:
        """Consume the next token and return it."""
        token = self.next_token
        assert token is not None

        self.position += len(token.text)
        self.next_token = None

        return token

    def raise_syntax_error(
        self,
        message: str,
        *,
        span_start: int | None = None,
        span_end: int | None = None,
    ) -> NoReturn:
        """Raise ParserSyntaxError at the given position."""
        span = (
            self.position if span_start is None else span_start,
            self.position if span_end is None else span_end,
        )
        raise ParserSyntaxError(
            message,
            source=self.source,
            span=span,
        )

    @contextlib.contextmanager
    def enclosing_tokens(
        self, open_token: str, close_token: str, *, around: str
    ) -> Iterator[None]:
        if self.check(open_token):
            open_position = self.position
            self.read()
        else:
            open_position = None

        yield

        if open_position is None:
            return

        if not self.check(close_token):
            self.raise_syntax_error(
                f"Expected matching {close_token} for {open_token}, after {around}",
                span_start=open_position,
            )

        self.read()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\packaging\_tokenizer.py ===
from __future__ import annotations

import contextlib
import re
from dataclasses import dataclass
from typing import Iterator, NoReturn

from .specifiers import Specifier


@dataclass
class Token:
    name: str
    text: str
    position: int


class ParserSyntaxError(Exception):
    """The provided source text could not be parsed correctly."""

    def __init__(
        self,
        message: str,
        *,
        source: str,
        span: tuple[int, int],
    ) -> None:
        self.span = span
        self.message = message
        self.source = source

        super().__init__()

    def __str__(self) -> str:
        marker = " " * self.span[0] + "~" * (self.span[1] - self.span[0]) + "^"
        return "\n    ".join([self.message, self.source, marker])


DEFAULT_RULES: dict[str, str | re.Pattern[str]] = {
    "LEFT_PARENTHESIS": r"\(",
    "RIGHT_PARENTHESIS": r"\)",
    "LEFT_BRACKET": r"\[",
    "RIGHT_BRACKET": r"\]",
    "SEMICOLON": r";",
    "COMMA": r",",
    "QUOTED_STRING": re.compile(
        r"""
            (
                ('[^']*')
                |
                ("[^"]*")
            )
        """,
        re.VERBOSE,
    ),
    "OP": r"(===|==|~=|!=|<=|>=|<|>)",
    "BOOLOP": r"\b(or|and)\b",
    "IN": r"\bin\b",
    "NOT": r"\bnot\b",
    "VARIABLE": re.compile(
        r"""
            \b(
                python_version
                |python_full_version
                |os[._]name
                |sys[._]platform
                |platform_(release|system)
                |platform[._](version|machine|python_implementation)
                |python_implementation
                |implementation_(name|version)
                |extras?
                |dependency_groups
            )\b
        """,
        re.VERBOSE,
    ),
    "SPECIFIER": re.compile(
        Specifier._operator_regex_str + Specifier._version_regex_str,
        re.VERBOSE | re.IGNORECASE,
    ),
    "AT": r"\@",
    "URL": r"[^ \t]+",
    "IDENTIFIER": r"\b[a-zA-Z0-9][a-zA-Z0-9._-]*\b",
    "VERSION_PREFIX_TRAIL": r"\.\*",
    "VERSION_LOCAL_LABEL_TRAIL": r"\+[a-z0-9]+(?:[-_\.][a-z0-9]+)*",
    "WS": r"[ \t]+",
    "END": r"$",
}


class Tokenizer:
    """Context-sensitive token parsing.

    Provides methods to examine the input stream to check whether the next token
    matches.
    """

    def __init__(
        self,
        source: str,
        *,
        rules: dict[str, str | re.Pattern[str]],
    ) -> None:
        self.source = source
        self.rules: dict[str, re.Pattern[str]] = {
            name: re.compile(pattern) for name, pattern in rules.items()
        }
        self.next_token: Token | None = None
        self.position = 0

    def consume(self, name: str) -> None:
        """Move beyond provided token name, if at current position."""
        if self.check(name):
            self.read()

    def check(self, name: str, *, peek: bool = False) -> bool:
        """Check whether the next token has the provided name.

        By default, if the check succeeds, the token *must* be read before
        another check. If `peek` is set to `True`, the token is not loaded and
        would need to be checked again.
        """
        assert self.next_token is None, (
            f"Cannot check for {name!r}, already have {self.next_token!r}"
        )
        assert name in self.rules, f"Unknown token name: {name!r}"

        expression = self.rules[name]

        match = expression.match(self.source, self.position)
        if match is None:
            return False
        if not peek:
            self.next_token = Token(name, match[0], self.position)
        return True

    def expect(self, name: str, *, expected: str) -> Token:
        """Expect a certain token name next, failing with a syntax error otherwise.

        The token is *not* read.
        """
        if not self.check(name):
            raise self.raise_syntax_error(f"Expected {expected}")
        return self.read()

    def read(self) -> Token:
        """Consume the next token and return it."""
        token = self.next_token
        assert token is not None

        self.position += len(token.text)
        self.next_token = None

        return token

    def raise_syntax_error(
        self,
        message: str,
        *,
        span_start: int | None = None,
        span_end: int | None = None,
    ) -> NoReturn:
        """Raise ParserSyntaxError at the given position."""
        span = (
            self.position if span_start is None else span_start,
            self.position if span_end is None else span_end,
        )
        raise ParserSyntaxError(
            message,
            source=self.source,
            span=span,
        )

    @contextlib.contextmanager
    def enclosing_tokens(
        self, open_token: str, close_token: str, *, around: str
    ) -> Iterator[None]:
        if self.check(open_token):
            open_position = self.position
            self.read()
        else:
            open_position = None

        yield

        if open_position is None:
            return

        if not self.check(close_token):
            self.raise_syntax_error(
                f"Expected matching {close_token} for {open_token}, after {around}",
                span_start=open_position,
            )

        self.read()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\google\protobuf\internal\extension_dict.py ===
# Protocol Buffers - Google's data interchange format
# Copyright 2008 Google Inc.  All rights reserved.
#
# Use of this source code is governed by a BSD-style
# license that can be found in the LICENSE file or at
# https://developers.google.com/open-source/licenses/bsd

"""Contains _ExtensionDict class to represent extensions.
"""

from google.protobuf.internal import type_checkers
from google.protobuf.descriptor import FieldDescriptor


def _VerifyExtensionHandle(message, extension_handle):
  """Verify that the given extension handle is valid."""

  if not isinstance(extension_handle, FieldDescriptor):
    raise KeyError('HasExtension() expects an extension handle, got: %s' %
                   extension_handle)

  if not extension_handle.is_extension:
    raise KeyError('"%s" is not an extension.' % extension_handle.full_name)

  if not extension_handle.containing_type:
    raise KeyError('"%s" is missing a containing_type.'
                   % extension_handle.full_name)

  if extension_handle.containing_type is not message.DESCRIPTOR:
    raise KeyError('Extension "%s" extends message type "%s", but this '
                   'message is of type "%s".' %
                   (extension_handle.full_name,
                    extension_handle.containing_type.full_name,
                    message.DESCRIPTOR.full_name))


# TODO: Unify error handling of "unknown extension" crap.
# TODO: Support iteritems()-style iteration over all
# extensions with the "has" bits turned on?
class _ExtensionDict(object):

  """Dict-like container for Extension fields on proto instances.

  Note that in all cases we expect extension handles to be
  FieldDescriptors.
  """

  def __init__(self, extended_message):
    """
    Args:
      extended_message: Message instance for which we are the Extensions dict.
    """
    self._extended_message = extended_message

  def __getitem__(self, extension_handle):
    """Returns the current value of the given extension handle."""

    _VerifyExtensionHandle(self._extended_message, extension_handle)

    result = self._extended_message._fields.get(extension_handle)
    if result is not None:
      return result

    if extension_handle.label == FieldDescriptor.LABEL_REPEATED:
      result = extension_handle._default_constructor(self._extended_message)
    elif extension_handle.cpp_type == FieldDescriptor.CPPTYPE_MESSAGE:
      message_type = extension_handle.message_type
      if not hasattr(message_type, '_concrete_class'):
        # pylint: disable=g-import-not-at-top
        from google.protobuf import message_factory
        message_factory.GetMessageClass(message_type)
      if not hasattr(extension_handle.message_type, '_concrete_class'):
        from google.protobuf import message_factory
        message_factory.GetMessageClass(extension_handle.message_type)
      result = extension_handle.message_type._concrete_class()
      try:
        result._SetListener(self._extended_message._listener_for_children)
      except ReferenceError:
        pass
    else:
      # Singular scalar -- just return the default without inserting into the
      # dict.
      return extension_handle.default_value

    # Atomically check if another thread has preempted us and, if not, swap
    # in the new object we just created.  If someone has preempted us, we
    # take that object and discard ours.
    # WARNING:  We are relying on setdefault() being atomic.  This is true
    #   in CPython but we haven't investigated others.  This warning appears
    #   in several other locations in this file.
    result = self._extended_message._fields.setdefault(
        extension_handle, result)

    return result

  def __eq__(self, other):
    if not isinstance(other, self.__class__):
      return False

    my_fields = self._extended_message.ListFields()
    other_fields = other._extended_message.ListFields()

    # Get rid of non-extension fields.
    my_fields = [field for field in my_fields if field.is_extension]
    other_fields = [field for field in other_fields if field.is_extension]

    return my_fields == other_fields

  def __ne__(self, other):
    return not self == other

  def __len__(self):
    fields = self._extended_message.ListFields()
    # Get rid of non-extension fields.
    extension_fields = [field for field in fields if field[0].is_extension]
    return len(extension_fields)

  def __hash__(self):
    raise TypeError('unhashable object')

  # Note that this is only meaningful for non-repeated, scalar extension
  # fields.  Note also that we may have to call _Modified() when we do
  # successfully set a field this way, to set any necessary "has" bits in the
  # ancestors of the extended message.
  def __setitem__(self, extension_handle, value):
    """If extension_handle specifies a non-repeated, scalar extension
    field, sets the value of that field.
    """

    _VerifyExtensionHandle(self._extended_message, extension_handle)

    if (extension_handle.label == FieldDescriptor.LABEL_REPEATED or
        extension_handle.cpp_type == FieldDescriptor.CPPTYPE_MESSAGE):
      raise TypeError(
          'Cannot assign to extension "%s" because it is a repeated or '
          'composite type.' % extension_handle.full_name)

    # It's slightly wasteful to lookup the type checker each time,
    # but we expect this to be a vanishingly uncommon case anyway.
    type_checker = type_checkers.GetTypeChecker(extension_handle)
    # pylint: disable=protected-access
    self._extended_message._fields[extension_handle] = (
        type_checker.CheckValue(value))
    self._extended_message._Modified()

  def __delitem__(self, extension_handle):
    self._extended_message.ClearExtension(extension_handle)

  def _FindExtensionByName(self, name):
    """Tries to find a known extension with the specified name.

    Args:
      name: Extension full name.

    Returns:
      Extension field descriptor.
    """
    descriptor = self._extended_message.DESCRIPTOR
    extensions = descriptor.file.pool._extensions_by_name[descriptor]
    return extensions.get(name, None)

  def _FindExtensionByNumber(self, number):
    """Tries to find a known extension with the field number.

    Args:
      number: Extension field number.

    Returns:
      Extension field descriptor.
    """
    descriptor = self._extended_message.DESCRIPTOR
    extensions = descriptor.file.pool._extensions_by_number[descriptor]
    return extensions.get(number, None)

  def __iter__(self):
    # Return a generator over the populated extension fields
    return (f[0] for f in self._extended_message.ListFields()
            if f[0].is_extension)

  def __contains__(self, extension_handle):
    _VerifyExtensionHandle(self._extended_message, extension_handle)

    if extension_handle not in self._extended_message._fields:
      return False

    if extension_handle.label == FieldDescriptor.LABEL_REPEATED:
      return bool(self._extended_message._fields.get(extension_handle))

    if extension_handle.cpp_type == FieldDescriptor.CPPTYPE_MESSAGE:
      value = self._extended_message._fields.get(extension_handle)
      # pylint: disable=protected-access
      return value is not None and value._is_present_in_parent

    return True

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\packaging\manifest.py ===
# Copyright (c) 2010-2024 openpyxl

"""
File manifest
"""
from mimetypes import MimeTypes
import os.path

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import String, Sequence
from openpyxl.xml.functions import fromstring
from openpyxl.xml.constants import (
    ARC_CONTENT_TYPES,
    ARC_THEME,
    ARC_STYLE,
    THEME_TYPE,
    STYLES_TYPE,
    CONTYPES_NS,
    ACTIVEX,
    CTRL,
    VBA,
)
from openpyxl.xml.functions import tostring

# initialise mime-types
mimetypes = MimeTypes()
mimetypes.add_type('application/xml', ".xml")
mimetypes.add_type('application/vnd.openxmlformats-package.relationships+xml', ".rels")
mimetypes.add_type("application/vnd.ms-office.vbaProject", ".bin")
mimetypes.add_type("application/vnd.openxmlformats-officedocument.vmlDrawing", ".vml")
mimetypes.add_type("image/x-emf", ".emf")


class FileExtension(Serialisable):

    tagname = "Default"

    Extension = String()
    ContentType = String()

    def __init__(self, Extension, ContentType):
        self.Extension = Extension
        self.ContentType = ContentType


class Override(Serialisable):

    tagname = "Override"

    PartName = String()
    ContentType = String()

    def __init__(self, PartName, ContentType):
        self.PartName = PartName
        self.ContentType = ContentType


DEFAULT_TYPES = [
    FileExtension("rels", "application/vnd.openxmlformats-package.relationships+xml"),
    FileExtension("xml", "application/xml"),
]

DEFAULT_OVERRIDE = [
    Override("/" + ARC_STYLE, STYLES_TYPE), # Styles
    Override("/" + ARC_THEME, THEME_TYPE), # Theme
    Override("/docProps/core.xml", "application/vnd.openxmlformats-package.core-properties+xml"),
    Override("/docProps/app.xml", "application/vnd.openxmlformats-officedocument.extended-properties+xml")
]


class Manifest(Serialisable):

    tagname = "Types"

    Default = Sequence(expected_type=FileExtension, unique=True)
    Override = Sequence(expected_type=Override, unique=True)
    path = "[Content_Types].xml"

    __elements__ = ("Default", "Override")

    def __init__(self,
                 Default=(),
                 Override=(),
                 ):
        if not Default:
            Default = DEFAULT_TYPES
        self.Default = Default
        if not Override:
            Override = DEFAULT_OVERRIDE
        self.Override = Override


    @property
    def filenames(self):
        return [part.PartName for part in self.Override]


    @property
    def extensions(self):
        """
        Map content types to file extensions
        Skip parts without extensions
        """
        exts = {os.path.splitext(part.PartName)[-1] for part in self.Override}
        return [(ext[1:], mimetypes.types_map[True][ext]) for ext in sorted(exts) if ext]


    def to_tree(self):
        """
        Custom serialisation method to allow setting a default namespace
        """
        defaults = [t.Extension for t in self.Default]
        for ext, mime in self.extensions:
            if ext not in defaults:
                mime = FileExtension(ext, mime)
                self.Default.append(mime)
        tree = super().to_tree()
        tree.set("xmlns", CONTYPES_NS)
        return tree


    def __contains__(self, content_type):
        """
        Check whether a particular content type is contained
        """
        for t in self.Override:
            if t.ContentType == content_type:
                return True


    def find(self, content_type):
        """
        Find specific content-type
        """
        try:
            return next(self.findall(content_type))
        except StopIteration:
            return


    def findall(self, content_type):
        """
        Find all elements of a specific content-type
        """
        for t in self.Override:
            if t.ContentType == content_type:
                yield t


    def append(self, obj):
        """
        Add content object to the package manifest
        # needs a contract...
        """
        ct = Override(PartName=obj.path, ContentType=obj.mime_type)
        self.Override.append(ct)


    def _write(self, archive, workbook):
        """
        Write manifest to the archive
        """
        self.append(workbook)
        self._write_vba(workbook)
        self._register_mimetypes(filenames=archive.namelist())
        archive.writestr(self.path, tostring(self.to_tree()))


    def _register_mimetypes(self, filenames):
        """
        Make sure that the mime type for all file extensions is registered
        """
        for fn in filenames:
            ext = os.path.splitext(fn)[-1]
            if not ext:
                continue
            mime = mimetypes.types_map[True][ext]
            fe = FileExtension(ext[1:], mime)
            self.Default.append(fe)


    def _write_vba(self, workbook):
        """
        Add content types from cached workbook when keeping VBA
        """
        if workbook.vba_archive:
            node = fromstring(workbook.vba_archive.read(ARC_CONTENT_TYPES))
            mf = Manifest.from_tree(node)
            filenames = self.filenames
            for override in mf.Override:
                if override.PartName not in (ACTIVEX, CTRL, VBA):
                    continue
                if override.PartName not in filenames:
                    self.Override.append(override)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\tensor\array\expressions\arrayexpr_derivatives.py ===
import operator
from functools import reduce, singledispatch

from sympy.core.expr import Expr
from sympy.core.singleton import S
from sympy.matrices.expressions.hadamard import HadamardProduct
from sympy.matrices.expressions.inverse import Inverse
from sympy.matrices.expressions.matexpr import (MatrixExpr, MatrixSymbol)
from sympy.matrices.expressions.special import Identity, OneMatrix
from sympy.matrices.expressions.transpose import Transpose
from sympy.combinatorics.permutations import _af_invert
from sympy.matrices.expressions.applyfunc import ElementwiseApplyFunction
from sympy.tensor.array.expressions.array_expressions import (
    _ArrayExpr, ZeroArray, ArraySymbol, ArrayTensorProduct, ArrayAdd,
    PermuteDims, ArrayDiagonal, ArrayElementwiseApplyFunc, get_rank,
    get_shape, ArrayContraction, _array_tensor_product, _array_contraction,
    _array_diagonal, _array_add, _permute_dims, Reshape)
from sympy.tensor.array.expressions.from_matrix_to_array import convert_matrix_to_array


@singledispatch
def array_derive(expr, x):
    """
    Derivatives (gradients) for array expressions.
    """
    raise NotImplementedError(f"not implemented for type {type(expr)}")


@array_derive.register(Expr)
def _(expr: Expr, x: _ArrayExpr):
    return ZeroArray(*x.shape)


@array_derive.register(ArrayTensorProduct)
def _(expr: ArrayTensorProduct, x: Expr):
    args = expr.args
    addend_list = []
    for i, arg in enumerate(expr.args):
        darg = array_derive(arg, x)
        if darg == 0:
            continue
        args_prev = args[:i]
        args_succ = args[i+1:]
        shape_prev = reduce(operator.add, map(get_shape, args_prev), ())
        shape_succ = reduce(operator.add, map(get_shape, args_succ), ())
        addend = _array_tensor_product(*args_prev, darg, *args_succ)
        tot1 = len(get_shape(x))
        tot2 = tot1 + len(shape_prev)
        tot3 = tot2 + len(get_shape(arg))
        tot4 = tot3 + len(shape_succ)
        perm = list(range(tot1, tot2)) + \
               list(range(tot1)) + list(range(tot2, tot3)) + \
               list(range(tot3, tot4))
        addend = _permute_dims(addend, _af_invert(perm))
        addend_list.append(addend)
    if len(addend_list) == 1:
        return addend_list[0]
    elif len(addend_list) == 0:
        return S.Zero
    else:
        return _array_add(*addend_list)


@array_derive.register(ArraySymbol)
def _(expr: ArraySymbol, x: _ArrayExpr):
    if expr == x:
        return _permute_dims(
            ArrayTensorProduct.fromiter(Identity(i) for i in expr.shape),
            [2*i for i in range(len(expr.shape))] + [2*i+1 for i in range(len(expr.shape))]
        )
    return ZeroArray(*(x.shape + expr.shape))


@array_derive.register(MatrixSymbol)
def _(expr: MatrixSymbol, x: _ArrayExpr):
    m, n = expr.shape
    if expr == x:
        return _permute_dims(
            _array_tensor_product(Identity(m), Identity(n)),
            [0, 2, 1, 3]
        )
    return ZeroArray(*(x.shape + expr.shape))


@array_derive.register(Identity)
def _(expr: Identity, x: _ArrayExpr):
    return ZeroArray(*(x.shape + expr.shape))


@array_derive.register(OneMatrix)
def _(expr: OneMatrix, x: _ArrayExpr):
    return ZeroArray(*(x.shape + expr.shape))


@array_derive.register(Transpose)
def _(expr: Transpose, x: Expr):
    # D(A.T, A) ==> (m,n,i,j) ==> D(A_ji, A_mn) = d_mj d_ni
    # D(B.T, A) ==> (m,n,i,j) ==> D(B_ji, A_mn)
    fd = array_derive(expr.arg, x)
    return _permute_dims(fd, [0, 1, 3, 2])


@array_derive.register(Inverse)
def _(expr: Inverse, x: Expr):
    mat = expr.I
    dexpr = array_derive(mat, x)
    tp = _array_tensor_product(-expr, dexpr, expr)
    mp = _array_contraction(tp, (1, 4), (5, 6))
    pp = _permute_dims(mp, [1, 2, 0, 3])
    return pp


@array_derive.register(ElementwiseApplyFunction)
def _(expr: ElementwiseApplyFunction, x: Expr):
    assert get_rank(expr) == 2
    assert get_rank(x) == 2
    fdiff = expr._get_function_fdiff()
    dexpr = array_derive(expr.expr, x)
    tp = _array_tensor_product(
        ElementwiseApplyFunction(fdiff, expr.expr),
        dexpr
    )
    td = _array_diagonal(
        tp, (0, 4), (1, 5)
    )
    return td


@array_derive.register(ArrayElementwiseApplyFunc)
def _(expr: ArrayElementwiseApplyFunc, x: Expr):
    fdiff = expr._get_function_fdiff()
    subexpr = expr.expr
    dsubexpr = array_derive(subexpr, x)
    tp = _array_tensor_product(
        dsubexpr,
        ArrayElementwiseApplyFunc(fdiff, subexpr)
    )
    b = get_rank(x)
    c = get_rank(expr)
    diag_indices = [(b + i, b + c + i) for i in range(c)]
    return _array_diagonal(tp, *diag_indices)


@array_derive.register(MatrixExpr)
def _(expr: MatrixExpr, x: Expr):
    cg = convert_matrix_to_array(expr)
    return array_derive(cg, x)


@array_derive.register(HadamardProduct)
def _(expr: HadamardProduct, x: Expr):
    raise NotImplementedError()


@array_derive.register(ArrayContraction)
def _(expr: ArrayContraction, x: Expr):
    fd = array_derive(expr.expr, x)
    rank_x = len(get_shape(x))
    contraction_indices = expr.contraction_indices
    new_contraction_indices = [tuple(j + rank_x for j in i) for i in contraction_indices]
    return _array_contraction(fd, *new_contraction_indices)


@array_derive.register(ArrayDiagonal)
def _(expr: ArrayDiagonal, x: Expr):
    dsubexpr = array_derive(expr.expr, x)
    rank_x = len(get_shape(x))
    diag_indices = [[j + rank_x for j in i] for i in expr.diagonal_indices]
    return _array_diagonal(dsubexpr, *diag_indices)


@array_derive.register(ArrayAdd)
def _(expr: ArrayAdd, x: Expr):
    return _array_add(*[array_derive(arg, x) for arg in expr.args])


@array_derive.register(PermuteDims)
def _(expr: PermuteDims, x: Expr):
    de = array_derive(expr.expr, x)
    perm = [0, 1] + [i + 2 for i in expr.permutation.array_form]
    return _permute_dims(de, perm)


@array_derive.register(Reshape)
def _(expr: Reshape, x: Expr):
    de = array_derive(expr.expr, x)
    return Reshape(de, get_shape(x) + expr.shape)


def matrix_derive(expr, x):
    from sympy.tensor.array.expressions.from_array_to_matrix import convert_array_to_matrix
    ce = convert_matrix_to_array(expr)
    dce = array_derive(ce, x)
    return convert_array_to_matrix(dce).doit()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\werkzeug\testapp.py ===
"""A small application that can be used to test a WSGI server and check
it for WSGI compliance.
"""

from __future__ import annotations

import importlib.metadata
import os
import sys
import typing as t
from textwrap import wrap

from markupsafe import escape

from .wrappers.request import Request
from .wrappers.response import Response

TEMPLATE = """\
<!doctype html>
<html lang=en>
<title>WSGI Information</title>
<style type="text/css">
  @import url(https://fonts.googleapis.com/css?family=Ubuntu);

  body       { font-family: 'Lucida Grande', 'Lucida Sans Unicode', 'Geneva',
               'Verdana', sans-serif; background-color: white; color: #000;
               font-size: 15px; text-align: center; }
  div.box    { text-align: left; width: 45em; margin: auto; padding: 50px 0;
               background-color: white; }
  h1, h2     { font-family: 'Ubuntu', 'Lucida Grande', 'Lucida Sans Unicode',
               'Geneva', 'Verdana', sans-serif; font-weight: normal; }
  h1         { margin: 0 0 30px 0; }
  h2         { font-size: 1.4em; margin: 1em 0 0.5em 0; }
  table      { width: 100%%; border-collapse: collapse; border: 1px solid #AFC5C9 }
  table th   { background-color: #AFC1C4; color: white; font-size: 0.72em;
               font-weight: normal; width: 18em; vertical-align: top;
               padding: 0.5em 0 0.1em 0.5em; }
  table td   { border: 1px solid #AFC5C9; padding: 0.1em 0 0.1em 0.5em; }
  code       { font-family: 'Consolas', 'Monaco', 'Bitstream Vera Sans Mono',
               monospace; font-size: 0.7em; }
  ul li      { line-height: 1.5em; }
  ul.path    { font-size: 0.7em; margin: 0 -30px; padding: 8px 30px;
               list-style: none; background: #E8EFF0; }
  ul.path li { line-height: 1.6em; }
  li.virtual { color: #999; text-decoration: underline; }
  li.exp     { background: white; }
</style>
<div class="box">
  <h1>WSGI Information</h1>
  <p>
    This page displays all available information about the WSGI server and
    the underlying Python interpreter.
  <h2 id="python-interpreter">Python Interpreter</h2>
  <table>
    <tr>
      <th>Python Version
      <td>%(python_version)s
    <tr>
      <th>Platform
      <td>%(platform)s [%(os)s]
    <tr>
      <th>API Version
      <td>%(api_version)s
    <tr>
      <th>Byteorder
      <td>%(byteorder)s
    <tr>
      <th>Werkzeug Version
      <td>%(werkzeug_version)s
  </table>
  <h2 id="wsgi-environment">WSGI Environment</h2>
  <table>%(wsgi_env)s</table>
  <h2 id="installed-eggs">Installed Eggs</h2>
  <p>
    The following python packages were installed on the system as
    Python eggs:
  <ul>%(python_eggs)s</ul>
  <h2 id="sys-path">System Path</h2>
  <p>
    The following paths are the current contents of the load path.  The
    following entries are looked up for Python packages.  Note that not
    all items in this path are folders.  Gray and underlined items are
    entries pointing to invalid resources or used by custom import hooks
    such as the zip importer.
  <p>
    Items with a bright background were expanded for display from a relative
    path.  If you encounter such paths in the output you might want to check
    your setup as relative paths are usually problematic in multithreaded
    environments.
  <ul class="path">%(sys_path)s</ul>
</div>
"""


def iter_sys_path() -> t.Iterator[tuple[str, bool, bool]]:
    if os.name == "posix":

        def strip(x: str) -> str:
            prefix = os.path.expanduser("~")
            if x.startswith(prefix):
                x = f"~{x[len(prefix) :]}"
            return x

    else:

        def strip(x: str) -> str:
            return x

    cwd = os.path.abspath(os.getcwd())
    for item in sys.path:
        path = os.path.join(cwd, item or os.path.curdir)
        yield strip(os.path.normpath(path)), not os.path.isdir(path), path != item


@Request.application
def test_app(req: Request) -> Response:
    """Simple test application that dumps the environment.  You can use
    it to check if Werkzeug is working properly:

    .. sourcecode:: pycon

        >>> from werkzeug.serving import run_simple
        >>> from werkzeug.testapp import test_app
        >>> run_simple('localhost', 3000, test_app)
         * Running on http://localhost:3000/

    The application displays important information from the WSGI environment,
    the Python interpreter and the installed libraries.
    """
    try:
        import pkg_resources
    except ImportError:
        eggs: t.Iterable[t.Any] = ()
    else:
        eggs = sorted(
            pkg_resources.working_set,
            key=lambda x: x.project_name.lower(),
        )
    python_eggs = []
    for egg in eggs:
        try:
            version = egg.version
        except (ValueError, AttributeError):
            version = "unknown"
        python_eggs.append(
            f"<li>{escape(egg.project_name)} <small>[{escape(version)}]</small>"
        )

    wsgi_env = []
    sorted_environ = sorted(req.environ.items(), key=lambda x: repr(x[0]).lower())
    for key, value in sorted_environ:
        value = "".join(wrap(str(escape(repr(value)))))
        wsgi_env.append(f"<tr><th>{escape(key)}<td><code>{value}</code>")

    sys_path = []
    for item, virtual, expanded in iter_sys_path():
        css = []
        if virtual:
            css.append("virtual")
        if expanded:
            css.append("exp")
        class_str = f' class="{" ".join(css)}"' if css else ""
        sys_path.append(f"<li{class_str}>{escape(item)}")

    context = {
        "python_version": "<br>".join(escape(sys.version).splitlines()),
        "platform": escape(sys.platform),
        "os": escape(os.name),
        "api_version": sys.api_version,
        "byteorder": sys.byteorder,
        "werkzeug_version": _get_werkzeug_version(),
        "python_eggs": "\n".join(python_eggs),
        "wsgi_env": "\n".join(wsgi_env),
        "sys_path": "\n".join(sys_path),
    }
    return Response(TEMPLATE % context, mimetype="text/html")


_werkzeug_version = ""


def _get_werkzeug_version() -> str:
    global _werkzeug_version

    if not _werkzeug_version:
        _werkzeug_version = importlib.metadata.version("werkzeug")

    return _werkzeug_version


if __name__ == "__main__":
    from .serving import run_simple

    run_simple("localhost", 5000, test_app, use_reloader=True)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\default_styles.py ===
from typing import Dict

from .style import Style

DEFAULT_STYLES: Dict[str, Style] = {
    "none": Style.null(),
    "reset": Style(
        color="default",
        bgcolor="default",
        dim=False,
        bold=False,
        italic=False,
        underline=False,
        blink=False,
        blink2=False,
        reverse=False,
        conceal=False,
        strike=False,
    ),
    "dim": Style(dim=True),
    "bright": Style(dim=False),
    "bold": Style(bold=True),
    "strong": Style(bold=True),
    "code": Style(reverse=True, bold=True),
    "italic": Style(italic=True),
    "emphasize": Style(italic=True),
    "underline": Style(underline=True),
    "blink": Style(blink=True),
    "blink2": Style(blink2=True),
    "reverse": Style(reverse=True),
    "strike": Style(strike=True),
    "black": Style(color="black"),
    "red": Style(color="red"),
    "green": Style(color="green"),
    "yellow": Style(color="yellow"),
    "magenta": Style(color="magenta"),
    "cyan": Style(color="cyan"),
    "white": Style(color="white"),
    "inspect.attr": Style(color="yellow", italic=True),
    "inspect.attr.dunder": Style(color="yellow", italic=True, dim=True),
    "inspect.callable": Style(bold=True, color="red"),
    "inspect.async_def": Style(italic=True, color="bright_cyan"),
    "inspect.def": Style(italic=True, color="bright_cyan"),
    "inspect.class": Style(italic=True, color="bright_cyan"),
    "inspect.error": Style(bold=True, color="red"),
    "inspect.equals": Style(),
    "inspect.help": Style(color="cyan"),
    "inspect.doc": Style(dim=True),
    "inspect.value.border": Style(color="green"),
    "live.ellipsis": Style(bold=True, color="red"),
    "layout.tree.row": Style(dim=False, color="red"),
    "layout.tree.column": Style(dim=False, color="blue"),
    "logging.keyword": Style(bold=True, color="yellow"),
    "logging.level.notset": Style(dim=True),
    "logging.level.debug": Style(color="green"),
    "logging.level.info": Style(color="blue"),
    "logging.level.warning": Style(color="yellow"),
    "logging.level.error": Style(color="red", bold=True),
    "logging.level.critical": Style(color="red", bold=True, reverse=True),
    "log.level": Style.null(),
    "log.time": Style(color="cyan", dim=True),
    "log.message": Style.null(),
    "log.path": Style(dim=True),
    "repr.ellipsis": Style(color="yellow"),
    "repr.indent": Style(color="green", dim=True),
    "repr.error": Style(color="red", bold=True),
    "repr.str": Style(color="green", italic=False, bold=False),
    "repr.brace": Style(bold=True),
    "repr.comma": Style(bold=True),
    "repr.ipv4": Style(bold=True, color="bright_green"),
    "repr.ipv6": Style(bold=True, color="bright_green"),
    "repr.eui48": Style(bold=True, color="bright_green"),
    "repr.eui64": Style(bold=True, color="bright_green"),
    "repr.tag_start": Style(bold=True),
    "repr.tag_name": Style(color="bright_magenta", bold=True),
    "repr.tag_contents": Style(color="default"),
    "repr.tag_end": Style(bold=True),
    "repr.attrib_name": Style(color="yellow", italic=False),
    "repr.attrib_equal": Style(bold=True),
    "repr.attrib_value": Style(color="magenta", italic=False),
    "repr.number": Style(color="cyan", bold=True, italic=False),
    "repr.number_complex": Style(color="cyan", bold=True, italic=False),  # same
    "repr.bool_true": Style(color="bright_green", italic=True),
    "repr.bool_false": Style(color="bright_red", italic=True),
    "repr.none": Style(color="magenta", italic=True),
    "repr.url": Style(underline=True, color="bright_blue", italic=False, bold=False),
    "repr.uuid": Style(color="bright_yellow", bold=False),
    "repr.call": Style(color="magenta", bold=True),
    "repr.path": Style(color="magenta"),
    "repr.filename": Style(color="bright_magenta"),
    "rule.line": Style(color="bright_green"),
    "rule.text": Style.null(),
    "json.brace": Style(bold=True),
    "json.bool_true": Style(color="bright_green", italic=True),
    "json.bool_false": Style(color="bright_red", italic=True),
    "json.null": Style(color="magenta", italic=True),
    "json.number": Style(color="cyan", bold=True, italic=False),
    "json.str": Style(color="green", italic=False, bold=False),
    "json.key": Style(color="blue", bold=True),
    "prompt": Style.null(),
    "prompt.choices": Style(color="magenta", bold=True),
    "prompt.default": Style(color="cyan", bold=True),
    "prompt.invalid": Style(color="red"),
    "prompt.invalid.choice": Style(color="red"),
    "pretty": Style.null(),
    "scope.border": Style(color="blue"),
    "scope.key": Style(color="yellow", italic=True),
    "scope.key.special": Style(color="yellow", italic=True, dim=True),
    "scope.equals": Style(color="red"),
    "table.header": Style(bold=True),
    "table.footer": Style(bold=True),
    "table.cell": Style.null(),
    "table.title": Style(italic=True),
    "table.caption": Style(italic=True, dim=True),
    "traceback.error": Style(color="red", italic=True),
    "traceback.border.syntax_error": Style(color="bright_red"),
    "traceback.border": Style(color="red"),
    "traceback.text": Style.null(),
    "traceback.title": Style(color="red", bold=True),
    "traceback.exc_type": Style(color="bright_red", bold=True),
    "traceback.exc_value": Style.null(),
    "traceback.offset": Style(color="bright_red", bold=True),
    "traceback.error_range": Style(underline=True, bold=True),
    "traceback.note": Style(color="green", bold=True),
    "traceback.group.border": Style(color="magenta"),
    "bar.back": Style(color="grey23"),
    "bar.complete": Style(color="rgb(249,38,114)"),
    "bar.finished": Style(color="rgb(114,156,31)"),
    "bar.pulse": Style(color="rgb(249,38,114)"),
    "progress.description": Style.null(),
    "progress.filesize": Style(color="green"),
    "progress.filesize.total": Style(color="green"),
    "progress.download": Style(color="green"),
    "progress.elapsed": Style(color="yellow"),
    "progress.percentage": Style(color="magenta"),
    "progress.remaining": Style(color="cyan"),
    "progress.data.speed": Style(color="red"),
    "progress.spinner": Style(color="green"),
    "status.spinner": Style(color="green"),
    "tree": Style(),
    "tree.line": Style(),
    "markdown.paragraph": Style(),
    "markdown.text": Style(),
    "markdown.em": Style(italic=True),
    "markdown.emph": Style(italic=True),  # For commonmark backwards compatibility
    "markdown.strong": Style(bold=True),
    "markdown.code": Style(bold=True, color="cyan", bgcolor="black"),
    "markdown.code_block": Style(color="cyan", bgcolor="black"),
    "markdown.block_quote": Style(color="magenta"),
    "markdown.list": Style(color="cyan"),
    "markdown.item": Style(),
    "markdown.item.bullet": Style(color="yellow", bold=True),
    "markdown.item.number": Style(color="yellow", bold=True),
    "markdown.hr": Style(color="yellow"),
    "markdown.h1.border": Style(),
    "markdown.h1": Style(bold=True),
    "markdown.h2": Style(bold=True, underline=True),
    "markdown.h3": Style(bold=True),
    "markdown.h4": Style(bold=True, dim=True),
    "markdown.h5": Style(underline=True),
    "markdown.h6": Style(italic=True),
    "markdown.h7": Style(italic=True, dim=True),
    "markdown.link": Style(color="bright_blue"),
    "markdown.link_url": Style(color="blue", underline=True),
    "markdown.s": Style(strike=True),
    "iso8601.date": Style(color="blue"),
    "iso8601.time": Style(color="magenta"),
    "iso8601.timezone": Style(color="yellow"),
}


if __name__ == "__main__":  # pragma: no cover
    import argparse
    import io

    from pip._vendor.rich.console import Console
    from pip._vendor.rich.table import Table
    from pip._vendor.rich.text import Text

    parser = argparse.ArgumentParser()
    parser.add_argument("--html", action="store_true", help="Export as HTML table")
    args = parser.parse_args()
    html: bool = args.html
    console = Console(record=True, width=70, file=io.StringIO()) if html else Console()

    table = Table("Name", "Styling")

    for style_name, style in DEFAULT_STYLES.items():
        table.add_row(Text(style_name, style=style), str(style))

    console.print(table)
    if html:
        print(console.export_html(inline_styles=True))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_core\_tests\test_mock_clock.py ===
import time
from math import inf

import pytest

from trio import sleep

from ... import _core
from .. import wait_all_tasks_blocked
from .._mock_clock import MockClock
from .._run import GLOBAL_RUN_CONTEXT
from .tutil import slow


def test_mock_clock() -> None:
    REAL_NOW = 123.0
    c = MockClock()
    c._real_clock = lambda: REAL_NOW
    repr(c)  # smoke test
    assert c.rate == 0
    assert c.current_time() == 0
    c.jump(1.2)
    assert c.current_time() == 1.2
    with pytest.raises(ValueError, match=r"^time can't go backwards$"):
        c.jump(-1)
    assert c.current_time() == 1.2
    assert c.deadline_to_sleep_time(1.1) == 0
    assert c.deadline_to_sleep_time(1.2) == 0
    assert c.deadline_to_sleep_time(1.3) > 999999

    with pytest.raises(ValueError, match=r"^rate must be >= 0$"):
        c.rate = -1
    assert c.rate == 0

    c.rate = 2
    assert c.current_time() == 1.2
    REAL_NOW += 1
    assert c.current_time() == 3.2
    assert c.deadline_to_sleep_time(3.1) == 0
    assert c.deadline_to_sleep_time(3.2) == 0
    assert c.deadline_to_sleep_time(4.2) == 0.5

    c.rate = 0.5
    assert c.current_time() == 3.2
    assert c.deadline_to_sleep_time(3.1) == 0
    assert c.deadline_to_sleep_time(3.2) == 0
    assert c.deadline_to_sleep_time(4.2) == 2.0

    c.jump(0.8)
    assert c.current_time() == 4.0
    REAL_NOW += 1
    assert c.current_time() == 4.5

    c2 = MockClock(rate=3)
    assert c2.rate == 3
    assert c2.current_time() < 10


async def test_mock_clock_autojump(mock_clock: MockClock) -> None:
    assert mock_clock.autojump_threshold == inf

    mock_clock.autojump_threshold = 0
    assert mock_clock.autojump_threshold == 0

    real_start = time.perf_counter()

    virtual_start = _core.current_time()
    for i in range(10):
        print(f"sleeping {10 * i} seconds")
        await sleep(10 * i)
        print("woke up!")
        assert virtual_start + 10 * i == _core.current_time()
        virtual_start = _core.current_time()

    real_duration = time.perf_counter() - real_start
    print(f"Slept {10 * sum(range(10))} seconds in {real_duration} seconds")
    assert real_duration < 1

    mock_clock.autojump_threshold = 0.02
    t = _core.current_time()
    # this should wake up before the autojump threshold triggers, so time
    # shouldn't change
    await wait_all_tasks_blocked()
    assert t == _core.current_time()
    # this should too
    await wait_all_tasks_blocked(0.01)
    assert t == _core.current_time()

    # set up a situation where the autojump task is blocked for a long long
    # time, to make sure that cancel-and-adjust-threshold logic is working
    mock_clock.autojump_threshold = 10000
    await wait_all_tasks_blocked()
    mock_clock.autojump_threshold = 0
    # if the above line didn't take affect immediately, then this would be
    # bad:
    # ignore ASYNC116, not sleep_forever, trying to test a large but finite sleep
    await sleep(100000)  # noqa: ASYNC116


async def test_mock_clock_autojump_interference(mock_clock: MockClock) -> None:
    mock_clock.autojump_threshold = 0.02

    mock_clock2 = MockClock()
    # messing with the autojump threshold of a clock that isn't actually
    # installed in the run loop shouldn't do anything.
    mock_clock2.autojump_threshold = 0.01

    # if the autojump_threshold of 0.01 were in effect, then the next line
    # would block forever, as the autojump task kept waking up to try to
    # jump the clock.
    await wait_all_tasks_blocked(0.015)

    # but the 0.02 limit does apply
    # ignore ASYNC116, not sleep_forever, trying to test a large but finite sleep
    await sleep(100000)  # noqa: ASYNC116


def test_mock_clock_autojump_preset() -> None:
    # Check that we can set the autojump_threshold before the clock is
    # actually in use, and it gets picked up
    mock_clock = MockClock(autojump_threshold=0.1)
    mock_clock.autojump_threshold = 0.01
    real_start = time.perf_counter()
    _core.run(sleep, 10000, clock=mock_clock)
    assert time.perf_counter() - real_start < 1


async def test_mock_clock_autojump_0_and_wait_all_tasks_blocked_0(
    mock_clock: MockClock,
) -> None:
    # Checks that autojump_threshold=0 doesn't interfere with
    # calling wait_all_tasks_blocked with the default cushion=0.

    mock_clock.autojump_threshold = 0

    record = []

    async def sleeper() -> None:
        await sleep(100)
        record.append("yawn")

    async def waiter() -> None:
        await wait_all_tasks_blocked()
        record.append("waiter woke")
        await sleep(1000)
        record.append("waiter done")

    async with _core.open_nursery() as nursery:
        nursery.start_soon(sleeper)
        nursery.start_soon(waiter)

    assert record == ["waiter woke", "yawn", "waiter done"]


@slow
async def test_mock_clock_autojump_0_and_wait_all_tasks_blocked_nonzero(
    mock_clock: MockClock,
) -> None:
    # Checks that autojump_threshold=0 doesn't interfere with
    # calling wait_all_tasks_blocked with a non-zero cushion.

    mock_clock.autojump_threshold = 0

    record = []

    async def sleeper() -> None:
        await sleep(100)
        record.append("yawn")

    async def waiter() -> None:
        await wait_all_tasks_blocked(1)
        record.append("waiter done")

    async with _core.open_nursery() as nursery:
        nursery.start_soon(sleeper)
        nursery.start_soon(waiter)

    assert record == ["waiter done", "yawn"]


async def test_initialization_doesnt_mutate_runner() -> None:
    before = (
        GLOBAL_RUN_CONTEXT.runner.clock,
        GLOBAL_RUN_CONTEXT.runner.clock_autojump_threshold,
    )

    MockClock(autojump_threshold=2, rate=3)

    after = (
        GLOBAL_RUN_CONTEXT.runner.clock,
        GLOBAL_RUN_CONTEXT.runner.clock_autojump_threshold,
    )
    assert before == after

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_utils\_inspect.py ===
"""Subset of inspect module from upstream python

We use this instead of upstream because upstream inspect is slow to import, and
significantly contributes to numpy import times. Importing this copy has almost
no overhead.

"""
import types

__all__ = ['getargspec', 'formatargspec']

# ----------------------------------------------------------- type-checking
def ismethod(object):
    """Return true if the object is an instance method.

    Instance method objects provide these attributes:
        __doc__         documentation string
        __name__        name with which this method was defined
        im_class        class object in which this method belongs
        im_func         function object containing implementation of method
        im_self         instance to which this method is bound, or None

    """
    return isinstance(object, types.MethodType)

def isfunction(object):
    """Return true if the object is a user-defined function.

    Function objects provide these attributes:
        __doc__         documentation string
        __name__        name with which this function was defined
        func_code       code object containing compiled function bytecode
        func_defaults   tuple of any default values for arguments
        func_doc        (same as __doc__)
        func_globals    global namespace in which this function was defined
        func_name       (same as __name__)

    """
    return isinstance(object, types.FunctionType)

def iscode(object):
    """Return true if the object is a code object.

    Code objects provide these attributes:
        co_argcount     number of arguments (not including * or ** args)
        co_code         string of raw compiled bytecode
        co_consts       tuple of constants used in the bytecode
        co_filename     name of file in which this code object was created
        co_firstlineno  number of first line in Python source code
        co_flags        bitmap: 1=optimized | 2=newlocals | 4=*arg | 8=**arg
        co_lnotab       encoded mapping of line numbers to bytecode indices
        co_name         name with which this code object was defined
        co_names        tuple of names of local variables
        co_nlocals      number of local variables
        co_stacksize    virtual machine stack space required
        co_varnames     tuple of names of arguments and local variables

    """
    return isinstance(object, types.CodeType)


# ------------------------------------------------ argument list extraction
# These constants are from Python's compile.h.
CO_OPTIMIZED, CO_NEWLOCALS, CO_VARARGS, CO_VARKEYWORDS = 1, 2, 4, 8

def getargs(co):
    """Get information about the arguments accepted by a code object.

    Three things are returned: (args, varargs, varkw), where 'args' is
    a list of argument names (possibly containing nested lists), and
    'varargs' and 'varkw' are the names of the * and ** arguments or None.

    """

    if not iscode(co):
        raise TypeError('arg is not a code object')

    nargs = co.co_argcount
    names = co.co_varnames
    args = list(names[:nargs])

    # The following acrobatics are for anonymous (tuple) arguments.
    # Which we do not need to support, so remove to avoid importing
    # the dis module.
    for i in range(nargs):
        if args[i][:1] in ['', '.']:
            raise TypeError("tuple function arguments are not supported")
    varargs = None
    if co.co_flags & CO_VARARGS:
        varargs = co.co_varnames[nargs]
        nargs = nargs + 1
    varkw = None
    if co.co_flags & CO_VARKEYWORDS:
        varkw = co.co_varnames[nargs]
    return args, varargs, varkw

def getargspec(func):
    """Get the names and default values of a function's arguments.

    A tuple of four things is returned: (args, varargs, varkw, defaults).
    'args' is a list of the argument names (it may contain nested lists).
    'varargs' and 'varkw' are the names of the * and ** arguments or None.
    'defaults' is an n-tuple of the default values of the last n arguments.

    """

    if ismethod(func):
        func = func.__func__
    if not isfunction(func):
        raise TypeError('arg is not a Python function')
    args, varargs, varkw = getargs(func.__code__)
    return args, varargs, varkw, func.__defaults__

def getargvalues(frame):
    """Get information about arguments passed into a particular frame.

    A tuple of four things is returned: (args, varargs, varkw, locals).
    'args' is a list of the argument names (it may contain nested lists).
    'varargs' and 'varkw' are the names of the * and ** arguments or None.
    'locals' is the locals dictionary of the given frame.

    """
    args, varargs, varkw = getargs(frame.f_code)
    return args, varargs, varkw, frame.f_locals

def joinseq(seq):
    if len(seq) == 1:
        return '(' + seq[0] + ',)'
    else:
        return '(' + ', '.join(seq) + ')'

def strseq(object, convert, join=joinseq):
    """Recursively walk a sequence, stringifying each element.

    """
    if type(object) in [list, tuple]:
        return join([strseq(_o, convert, join) for _o in object])
    else:
        return convert(object)

def formatargspec(args, varargs=None, varkw=None, defaults=None,
                  formatarg=str,
                  formatvarargs=lambda name: '*' + name,
                  formatvarkw=lambda name: '**' + name,
                  formatvalue=lambda value: '=' + repr(value),
                  join=joinseq):
    """Format an argument spec from the 4 values returned by getargspec.

    The first four arguments are (args, varargs, varkw, defaults).  The
    other four arguments are the corresponding optional formatting functions
    that are called to turn names and values into strings.  The ninth
    argument is an optional function to format the sequence of arguments.

    """
    specs = []
    if defaults:
        firstdefault = len(args) - len(defaults)
    for i in range(len(args)):
        spec = strseq(args[i], formatarg, join)
        if defaults and i >= firstdefault:
            spec = spec + formatvalue(defaults[i - firstdefault])
        specs.append(spec)
    if varargs is not None:
        specs.append(formatvarargs(varargs))
    if varkw is not None:
        specs.append(formatvarkw(varkw))
    return '(' + ', '.join(specs) + ')'

def formatargvalues(args, varargs, varkw, locals,
                    formatarg=str,
                    formatvarargs=lambda name: '*' + name,
                    formatvarkw=lambda name: '**' + name,
                    formatvalue=lambda value: '=' + repr(value),
                    join=joinseq):
    """Format an argument spec from the 4 values returned by getargvalues.

    The first four arguments are (args, varargs, varkw, locals).  The
    next four arguments are the corresponding optional formatting functions
    that are called to turn names and values into strings.  The ninth
    argument is an optional function to format the sequence of arguments.

    """
    def convert(name, locals=locals,
                formatarg=formatarg, formatvalue=formatvalue):
        return formatarg(name) + formatvalue(locals[name])
    specs = [strseq(arg, convert, join) for arg in args]

    if varargs:
        specs.append(formatvarargs(varargs) + formatvalue(locals[varargs]))
    if varkw:
        specs.append(formatvarkw(varkw) + formatvalue(locals[varkw]))
    return '(' + ', '.join(specs) + ')'

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\quantization\shape_inference.py ===
# --------------------------------------------------------------------------
# Copyright (c) Microsoft, Intel Corporation. All rights reserved.
# Licensed under the MIT License. See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------


import logging
import tempfile
import traceback
from pathlib import Path

import onnx

import onnxruntime
from onnxruntime.tools.symbolic_shape_infer import SymbolicShapeInference
from onnxruntime.transformers.onnx_utils import extract_raw_data_from_model, has_external_data

from .quant_utils import add_pre_process_metadata

logger = logging.getLogger(__name__)


def quant_pre_process(
    input_model: str | Path | onnx.ModelProto | None = None,
    output_model_path: str | Path | None = None,
    skip_optimization: bool = False,
    skip_onnx_shape: bool = False,
    skip_symbolic_shape: bool = False,
    auto_merge: bool = False,
    int_max: int = 2**31 - 1,
    guess_output_rank: bool = False,
    verbose: int = 0,
    save_as_external_data: bool = False,
    all_tensors_to_one_file: bool = False,
    external_data_location: str | None = None,
    external_data_size_threshold: int = 1024,
    **deprecated_kwargs,
) -> None:
    """Shape inference and model optimization, in preparation for quantization.

    Args:
        input_model: Path to the input model file or ModelProto
        output_model_path: Path to the output model file
        skip_optimization: Skip model optimization step if true. This may result in ONNX shape
            inference failure for some models.
        skip_onnx_shape: Skip ONNX shape inference. Symbolic shape inference is most effective
            with transformer based models. Skipping all shape inferences may
            reduce the effectiveness of quantization, as a tensor with unknown
            shape can not be quantized.
        skip_symbolic_shape: Skip symbolic shape inference. Symbolic shape inference is most
            effective with transformer based models. Skipping all shape
            inferences may reduce the effectiveness of quantization, as a tensor
            with unknown shape can not be quantized.
        auto_merge: For symbolic shape inference, automatically merge symbolic dims when
            conflict happens.
        int_max: For symbolic shape inference, specify the maximum value for integer to be
            treated as boundless for ops like slice
        guess_output_rank: Guess output rank to be the same as input 0 for unknown ops
        verbose: Logs detailed info of inference, 0: turn off, 1: warnings, 3: detailed
        save_as_external_data: Saving an ONNX model to external data
        all_tensors_to_one_file: Saving all the external data to one file
        external_data_location: The file location to save the external file
        external_data_size_threshold: The size threshold for external data
    """

    if input_model is None:
        input_model = deprecated_kwargs.pop("input_model_path", None)
    assert input_model is not None

    assert output_model_path is not None, "output_model_path is required."

    with tempfile.TemporaryDirectory(prefix="pre.quant.") as quant_tmp_dir:
        temp_path = Path(quant_tmp_dir)
        model = None

        if not skip_symbolic_shape:
            logger.info("Performing symbolic shape inference...")
            loaded_model = input_model if isinstance(input_model, onnx.ModelProto) else onnx.load(input_model)
            model = SymbolicShapeInference.infer_shapes(
                loaded_model,
                int_max,
                auto_merge,
                guess_output_rank,
                verbose,
            )

        if not skip_optimization:
            # Use ORT optimizers (native code) to optimize model
            if not skip_symbolic_shape:
                # Need to save the inferenced model to file so as to run the optimizer
                input_model = str(temp_path / "symbolic_shape_inferred.onnx")
                if save_as_external_data:
                    onnx.save_model(
                        model,
                        input_model,
                        save_as_external_data=True,
                        all_tensors_to_one_file=all_tensors_to_one_file,
                        size_threshold=external_data_size_threshold,
                        convert_attribute=False,
                    )
                else:
                    onnx.save(model, input_model)
                model = None

            opt_model_path = str(temp_path / "optimized.onnx")
            try:
                sess_option = onnxruntime.SessionOptions()
                sess_option.optimized_model_filepath = opt_model_path
                sess_option.graph_optimization_level = onnxruntime.GraphOptimizationLevel.ORT_ENABLE_BASIC
                # For large model, extract external data from model and add to session options
                if isinstance(input_model, onnx.ModelProto):
                    if has_external_data(input_model):
                        raise ValueError(
                            "ModelProto has external data not loaded into memory, ORT cannot create session. "
                            "Please load external data before calling this function. "
                            "See https://onnx.ai/onnx/repo-docs/ExternalData.html for more information."
                        )
                    external_names, external_values = extract_raw_data_from_model(input_model)
                    sess_option.add_external_initializers(list(external_names), list(external_values))
                    input_model = input_model.SerializeToString()
                # the saved optimized model otherwise points to the original external data file name
                # which is not available relative to the optimized model file
                elif skip_symbolic_shape and save_as_external_data:
                    sess_option.add_session_config_entry(
                        "session.optimized_model_external_initializers_file_name", "optimized.onnx.data"
                    )

                sess = onnxruntime.InferenceSession(input_model, sess_option, providers=["CPUExecutionProvider"])
                # Close the session to avoid the cleanup error on Windows for temp folders
                # https://github.com/microsoft/onnxruntime/issues/17627
                del sess
            except Exception:
                logger.error(
                    "ONNX Runtime Model Optimization Failed! Consider rerun with option `--skip_optimization'."
                )
                logger.error(traceback.format_exc())

            input_model = opt_model_path

        if not skip_onnx_shape:
            # ONNX shape inference.
            # According to docs, infer_shapes_path should be used for 2G+ models.
            # If the skip optimization is specified, we could be dealing with a
            # large model. So be on the safe side, save the model
            if model is not None:
                input_model = str(temp_path / "symbolic_shape_inferred.onnx")
                if save_as_external_data:
                    onnx.save_model(
                        model,
                        input_model,
                        save_as_external_data=True,
                        all_tensors_to_one_file=all_tensors_to_one_file,
                        size_threshold=external_data_size_threshold,
                        convert_attribute=False,
                    )
                else:
                    onnx.save(model, input_model)
                model = None

            if isinstance(input_model, onnx.ModelProto):
                input_model = str(Path(quant_tmp_dir) / "model_input.onnx")
                onnx.save_model(
                    model,
                    input_model,
                    save_as_external_data=True,
                    all_tensors_to_one_file=all_tensors_to_one_file,
                    size_threshold=external_data_size_threshold,
                    convert_attribute=False,
                )

            inferred_model_path = str(temp_path / "onnx_shape_inferred.onnx")
            onnx.shape_inference.infer_shapes_path(input_model, inferred_model_path)
            model = onnx.load(inferred_model_path)

    if model is None:
        model = input_model if isinstance(input_model, onnx.ModelProto) else onnx.load(input_model)

    add_pre_process_metadata(model)

    if save_as_external_data:
        onnx.save_model(
            model,
            output_model_path,
            save_as_external_data=True,
            all_tensors_to_one_file=all_tensors_to_one_file,
            location=external_data_location,
            size_threshold=external_data_size_threshold,
            convert_attribute=False,
        )
    else:
        onnx.save(model, output_model_path)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\bidi\browser.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

from typing import Dict, List

from selenium.webdriver.common.bidi.common import command_builder


class ClientWindowState:
    """Represents a window state."""

    FULLSCREEN = "fullscreen"
    MAXIMIZED = "maximized"
    MINIMIZED = "minimized"
    NORMAL = "normal"


class ClientWindowInfo:
    """Represents a client window information."""

    def __init__(
        self,
        client_window: str,
        state: str,
        width: int,
        height: int,
        x: int,
        y: int,
        active: bool,
    ):
        self.client_window = client_window
        self.state = state
        self.width = width
        self.height = height
        self.x = x
        self.y = y
        self.active = active

    def get_state(self) -> str:
        """Gets the state of the client window.

        Returns:
        -------
            str: The state of the client window (one of the ClientWindowState constants).
        """
        return self.state

    def get_client_window(self) -> str:
        """Gets the client window identifier.

        Returns:
        -------
            str: The client window identifier.
        """
        return self.client_window

    def get_width(self) -> int:
        """Gets the width of the client window.

        Returns:
        -------
            int: The width of the client window.
        """
        return self.width

    def get_height(self) -> int:
        """Gets the height of the client window.

        Returns:
        -------
            int: The height of the client window.
        """
        return self.height

    def get_x(self) -> int:
        """Gets the x coordinate of the client window.

        Returns:
        -------
            int: The x coordinate of the client window.
        """
        return self.x

    def get_y(self) -> int:
        """Gets the y coordinate of the client window.

        Returns:
        -------
            int: The y coordinate of the client window.
        """
        return self.y

    def is_active(self) -> bool:
        """Checks if the client window is active.

        Returns:
        -------
            bool: True if the client window is active, False otherwise.
        """
        return self.active

    @classmethod
    def from_dict(cls, data: Dict) -> "ClientWindowInfo":
        """Creates a ClientWindowInfo instance from a dictionary.

        Parameters:
        -----------
            data: A dictionary containing the client window information.

        Returns:
        -------
            ClientWindowInfo: A new instance of ClientWindowInfo.
        """
        return cls(
            client_window=data.get("clientWindow"),
            state=data.get("state"),
            width=data.get("width"),
            height=data.get("height"),
            x=data.get("x"),
            y=data.get("y"),
            active=data.get("active"),
        )


class Browser:
    """
    BiDi implementation of the browser module.
    """

    def __init__(self, conn):
        self.conn = conn

    def create_user_context(self) -> str:
        """Creates a new user context.

        Returns:
        -------
            str: The ID of the created user context.
        """
        result = self.conn.execute(command_builder("browser.createUserContext", {}))
        return result["userContext"]

    def get_user_contexts(self) -> List[str]:
        """Gets all user contexts.

        Returns:
        -------
            List[str]: A list of user context IDs.
        """
        result = self.conn.execute(command_builder("browser.getUserContexts", {}))
        return [context_info["userContext"] for context_info in result["userContexts"]]

    def remove_user_context(self, user_context_id: str) -> None:
        """Removes a user context.

        Parameters:
        -----------
            user_context_id: The ID of the user context to remove.

        Raises:
        ------
            Exception: If the user context ID is "default" or does not exist.
        """
        if user_context_id == "default":
            raise Exception("Cannot remove the default user context")

        params = {"userContext": user_context_id}
        self.conn.execute(command_builder("browser.removeUserContext", params))

    def get_client_windows(self) -> List[ClientWindowInfo]:
        """Gets all client windows.

        Returns:
        -------
            List[ClientWindowInfo]: A list of client window information.
        """
        result = self.conn.execute(command_builder("browser.getClientWindows", {}))
        return [ClientWindowInfo.from_dict(window) for window in result["clientWindows"]]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\flask\views.py ===
from __future__ import annotations

import typing as t

from . import typing as ft
from .globals import current_app
from .globals import request

F = t.TypeVar("F", bound=t.Callable[..., t.Any])

http_method_funcs = frozenset(
    ["get", "post", "head", "options", "delete", "put", "trace", "patch"]
)


class View:
    """Subclass this class and override :meth:`dispatch_request` to
    create a generic class-based view. Call :meth:`as_view` to create a
    view function that creates an instance of the class with the given
    arguments and calls its ``dispatch_request`` method with any URL
    variables.

    See :doc:`views` for a detailed guide.

    .. code-block:: python

        class Hello(View):
            init_every_request = False

            def dispatch_request(self, name):
                return f"Hello, {name}!"

        app.add_url_rule(
            "/hello/<name>", view_func=Hello.as_view("hello")
        )

    Set :attr:`methods` on the class to change what methods the view
    accepts.

    Set :attr:`decorators` on the class to apply a list of decorators to
    the generated view function. Decorators applied to the class itself
    will not be applied to the generated view function!

    Set :attr:`init_every_request` to ``False`` for efficiency, unless
    you need to store request-global data on ``self``.
    """

    #: The methods this view is registered for. Uses the same default
    #: (``["GET", "HEAD", "OPTIONS"]``) as ``route`` and
    #: ``add_url_rule`` by default.
    methods: t.ClassVar[t.Collection[str] | None] = None

    #: Control whether the ``OPTIONS`` method is handled automatically.
    #: Uses the same default (``True``) as ``route`` and
    #: ``add_url_rule`` by default.
    provide_automatic_options: t.ClassVar[bool | None] = None

    #: A list of decorators to apply, in order, to the generated view
    #: function. Remember that ``@decorator`` syntax is applied bottom
    #: to top, so the first decorator in the list would be the bottom
    #: decorator.
    #:
    #: .. versionadded:: 0.8
    decorators: t.ClassVar[list[t.Callable[..., t.Any]]] = []

    #: Create a new instance of this view class for every request by
    #: default. If a view subclass sets this to ``False``, the same
    #: instance is used for every request.
    #:
    #: A single instance is more efficient, especially if complex setup
    #: is done during init. However, storing data on ``self`` is no
    #: longer safe across requests, and :data:`~flask.g` should be used
    #: instead.
    #:
    #: .. versionadded:: 2.2
    init_every_request: t.ClassVar[bool] = True

    def dispatch_request(self) -> ft.ResponseReturnValue:
        """The actual view function behavior. Subclasses must override
        this and return a valid response. Any variables from the URL
        rule are passed as keyword arguments.
        """
        raise NotImplementedError()

    @classmethod
    def as_view(
        cls, name: str, *class_args: t.Any, **class_kwargs: t.Any
    ) -> ft.RouteCallable:
        """Convert the class into a view function that can be registered
        for a route.

        By default, the generated view will create a new instance of the
        view class for every request and call its
        :meth:`dispatch_request` method. If the view class sets
        :attr:`init_every_request` to ``False``, the same instance will
        be used for every request.

        Except for ``name``, all other arguments passed to this method
        are forwarded to the view class ``__init__`` method.

        .. versionchanged:: 2.2
            Added the ``init_every_request`` class attribute.
        """
        if cls.init_every_request:

            def view(**kwargs: t.Any) -> ft.ResponseReturnValue:
                self = view.view_class(  # type: ignore[attr-defined]
                    *class_args, **class_kwargs
                )
                return current_app.ensure_sync(self.dispatch_request)(**kwargs)  # type: ignore[no-any-return]

        else:
            self = cls(*class_args, **class_kwargs)  # pyright: ignore

            def view(**kwargs: t.Any) -> ft.ResponseReturnValue:
                return current_app.ensure_sync(self.dispatch_request)(**kwargs)  # type: ignore[no-any-return]

        if cls.decorators:
            view.__name__ = name
            view.__module__ = cls.__module__
            for decorator in cls.decorators:
                view = decorator(view)

        # We attach the view class to the view function for two reasons:
        # first of all it allows us to easily figure out what class-based
        # view this thing came from, secondly it's also used for instantiating
        # the view class so you can actually replace it with something else
        # for testing purposes and debugging.
        view.view_class = cls  # type: ignore
        view.__name__ = name
        view.__doc__ = cls.__doc__
        view.__module__ = cls.__module__
        view.methods = cls.methods  # type: ignore
        view.provide_automatic_options = cls.provide_automatic_options  # type: ignore
        return view


class MethodView(View):
    """Dispatches request methods to the corresponding instance methods.
    For example, if you implement a ``get`` method, it will be used to
    handle ``GET`` requests.

    This can be useful for defining a REST API.

    :attr:`methods` is automatically set based on the methods defined on
    the class.

    See :doc:`views` for a detailed guide.

    .. code-block:: python

        class CounterAPI(MethodView):
            def get(self):
                return str(session.get("counter", 0))

            def post(self):
                session["counter"] = session.get("counter", 0) + 1
                return redirect(url_for("counter"))

        app.add_url_rule(
            "/counter", view_func=CounterAPI.as_view("counter")
        )
    """

    def __init_subclass__(cls, **kwargs: t.Any) -> None:
        super().__init_subclass__(**kwargs)

        if "methods" not in cls.__dict__:
            methods = set()

            for base in cls.__bases__:
                if getattr(base, "methods", None):
                    methods.update(base.methods)  # type: ignore[attr-defined]

            for key in http_method_funcs:
                if hasattr(cls, key):
                    methods.add(key.upper())

            if methods:
                cls.methods = methods

    def dispatch_request(self, **kwargs: t.Any) -> ft.ResponseReturnValue:
        meth = getattr(self, request.method.lower(), None)

        # If the request method is HEAD and we don't have a handler for it
        # retry with GET.
        if meth is None and request.method == "HEAD":
            meth = getattr(self, "get", None)

        assert meth is not None, f"Unimplemented method {request.method!r}"
        return current_app.ensure_sync(meth)(**kwargs)  # type: ignore[no-any-return]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\jinja2\debug.py ===
import sys
import typing as t
from types import CodeType
from types import TracebackType

from .exceptions import TemplateSyntaxError
from .utils import internal_code
from .utils import missing

if t.TYPE_CHECKING:
    from .runtime import Context


def rewrite_traceback_stack(source: t.Optional[str] = None) -> BaseException:
    """Rewrite the current exception to replace any tracebacks from
    within compiled template code with tracebacks that look like they
    came from the template source.

    This must be called within an ``except`` block.

    :param source: For ``TemplateSyntaxError``, the original source if
        known.
    :return: The original exception with the rewritten traceback.
    """
    _, exc_value, tb = sys.exc_info()
    exc_value = t.cast(BaseException, exc_value)
    tb = t.cast(TracebackType, tb)

    if isinstance(exc_value, TemplateSyntaxError) and not exc_value.translated:
        exc_value.translated = True
        exc_value.source = source
        # Remove the old traceback, otherwise the frames from the
        # compiler still show up.
        exc_value.with_traceback(None)
        # Outside of runtime, so the frame isn't executing template
        # code, but it still needs to point at the template.
        tb = fake_traceback(
            exc_value, None, exc_value.filename or "<unknown>", exc_value.lineno
        )
    else:
        # Skip the frame for the render function.
        tb = tb.tb_next

    stack = []

    # Build the stack of traceback object, replacing any in template
    # code with the source file and line information.
    while tb is not None:
        # Skip frames decorated with @internalcode. These are internal
        # calls that aren't useful in template debugging output.
        if tb.tb_frame.f_code in internal_code:
            tb = tb.tb_next
            continue

        template = tb.tb_frame.f_globals.get("__jinja_template__")

        if template is not None:
            lineno = template.get_corresponding_lineno(tb.tb_lineno)
            fake_tb = fake_traceback(exc_value, tb, template.filename, lineno)
            stack.append(fake_tb)
        else:
            stack.append(tb)

        tb = tb.tb_next

    tb_next = None

    # Assign tb_next in reverse to avoid circular references.
    for tb in reversed(stack):
        tb.tb_next = tb_next
        tb_next = tb

    return exc_value.with_traceback(tb_next)


def fake_traceback(  # type: ignore
    exc_value: BaseException, tb: t.Optional[TracebackType], filename: str, lineno: int
) -> TracebackType:
    """Produce a new traceback object that looks like it came from the
    template source instead of the compiled code. The filename, line
    number, and location name will point to the template, and the local
    variables will be the current template context.

    :param exc_value: The original exception to be re-raised to create
        the new traceback.
    :param tb: The original traceback to get the local variables and
        code info from.
    :param filename: The template filename.
    :param lineno: The line number in the template source.
    """
    if tb is not None:
        # Replace the real locals with the context that would be
        # available at that point in the template.
        locals = get_template_locals(tb.tb_frame.f_locals)
        locals.pop("__jinja_exception__", None)
    else:
        locals = {}

    globals = {
        "__name__": filename,
        "__file__": filename,
        "__jinja_exception__": exc_value,
    }
    # Raise an exception at the correct line number.
    code: CodeType = compile(
        "\n" * (lineno - 1) + "raise __jinja_exception__", filename, "exec"
    )

    # Build a new code object that points to the template file and
    # replaces the location with a block name.
    location = "template"

    if tb is not None:
        function = tb.tb_frame.f_code.co_name

        if function == "root":
            location = "top-level template code"
        elif function.startswith("block_"):
            location = f"block {function[6:]!r}"

    if sys.version_info >= (3, 8):
        code = code.replace(co_name=location)
    else:
        code = CodeType(
            code.co_argcount,
            code.co_kwonlyargcount,
            code.co_nlocals,
            code.co_stacksize,
            code.co_flags,
            code.co_code,
            code.co_consts,
            code.co_names,
            code.co_varnames,
            code.co_filename,
            location,
            code.co_firstlineno,
            code.co_lnotab,
            code.co_freevars,
            code.co_cellvars,
        )

    # Execute the new code, which is guaranteed to raise, and return
    # the new traceback without this frame.
    try:
        exec(code, globals, locals)
    except BaseException:
        return sys.exc_info()[2].tb_next  # type: ignore


def get_template_locals(real_locals: t.Mapping[str, t.Any]) -> t.Dict[str, t.Any]:
    """Based on the runtime locals, get the context that would be
    available at that point in the template.
    """
    # Start with the current template context.
    ctx: t.Optional[Context] = real_locals.get("context")

    if ctx is not None:
        data: t.Dict[str, t.Any] = ctx.get_all().copy()
    else:
        data = {}

    # Might be in a derived context that only sets local variables
    # rather than pushing a context. Local variables follow the scheme
    # l_depth_name. Find the highest-depth local that has a value for
    # each name.
    local_overrides: t.Dict[str, t.Tuple[int, t.Any]] = {}

    for name, value in real_locals.items():
        if not name.startswith("l_") or value is missing:
            # Not a template variable, or no longer relevant.
            continue

        try:
            _, depth_str, name = name.split("_", 2)
            depth = int(depth_str)
        except ValueError:
            continue

        cur_depth = local_overrides.get(name, (-1,))[0]

        if cur_depth < depth:
            local_overrides[name] = (depth, value)

    # Modify the context with any derived context.
    for name, (_, value) in local_overrides.items():
        if value is missing:
            data.pop(name, None)
        else:
            data[name] = value

    return data

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\urllib3\request.py ===
from __future__ import absolute_import

import sys

from .filepost import encode_multipart_formdata
from .packages import six
from .packages.six.moves.urllib.parse import urlencode

__all__ = ["RequestMethods"]


class RequestMethods(object):
    """
    Convenience mixin for classes who implement a :meth:`urlopen` method, such
    as :class:`urllib3.HTTPConnectionPool` and
    :class:`urllib3.PoolManager`.

    Provides behavior for making common types of HTTP request methods and
    decides which type of request field encoding to use.

    Specifically,

    :meth:`.request_encode_url` is for sending requests whose fields are
    encoded in the URL (such as GET, HEAD, DELETE).

    :meth:`.request_encode_body` is for sending requests whose fields are
    encoded in the *body* of the request using multipart or www-form-urlencoded
    (such as for POST, PUT, PATCH).

    :meth:`.request` is for making any kind of request, it will look up the
    appropriate encoding format and use one of the above two methods to make
    the request.

    Initializer parameters:

    :param headers:
        Headers to include with all requests, unless other headers are given
        explicitly.
    """

    _encode_url_methods = {"DELETE", "GET", "HEAD", "OPTIONS"}

    def __init__(self, headers=None):
        self.headers = headers or {}

    def urlopen(
        self,
        method,
        url,
        body=None,
        headers=None,
        encode_multipart=True,
        multipart_boundary=None,
        **kw
    ):  # Abstract
        raise NotImplementedError(
            "Classes extending RequestMethods must implement "
            "their own ``urlopen`` method."
        )

    def request(self, method, url, fields=None, headers=None, **urlopen_kw):
        """
        Make a request using :meth:`urlopen` with the appropriate encoding of
        ``fields`` based on the ``method`` used.

        This is a convenience method that requires the least amount of manual
        effort. It can be used in most situations, while still having the
        option to drop down to more specific methods when necessary, such as
        :meth:`request_encode_url`, :meth:`request_encode_body`,
        or even the lowest level :meth:`urlopen`.
        """
        method = method.upper()

        urlopen_kw["request_url"] = url

        if method in self._encode_url_methods:
            return self.request_encode_url(
                method, url, fields=fields, headers=headers, **urlopen_kw
            )
        else:
            return self.request_encode_body(
                method, url, fields=fields, headers=headers, **urlopen_kw
            )

    def request_encode_url(self, method, url, fields=None, headers=None, **urlopen_kw):
        """
        Make a request using :meth:`urlopen` with the ``fields`` encoded in
        the url. This is useful for request methods like GET, HEAD, DELETE, etc.
        """
        if headers is None:
            headers = self.headers

        extra_kw = {"headers": headers}
        extra_kw.update(urlopen_kw)

        if fields:
            url += "?" + urlencode(fields)

        return self.urlopen(method, url, **extra_kw)

    def request_encode_body(
        self,
        method,
        url,
        fields=None,
        headers=None,
        encode_multipart=True,
        multipart_boundary=None,
        **urlopen_kw
    ):
        """
        Make a request using :meth:`urlopen` with the ``fields`` encoded in
        the body. This is useful for request methods like POST, PUT, PATCH, etc.

        When ``encode_multipart=True`` (default), then
        :func:`urllib3.encode_multipart_formdata` is used to encode
        the payload with the appropriate content type. Otherwise
        :func:`urllib.parse.urlencode` is used with the
        'application/x-www-form-urlencoded' content type.

        Multipart encoding must be used when posting files, and it's reasonably
        safe to use it in other times too. However, it may break request
        signing, such as with OAuth.

        Supports an optional ``fields`` parameter of key/value strings AND
        key/filetuple. A filetuple is a (filename, data, MIME type) tuple where
        the MIME type is optional. For example::

            fields = {
                'foo': 'bar',
                'fakefile': ('foofile.txt', 'contents of foofile'),
                'realfile': ('barfile.txt', open('realfile').read()),
                'typedfile': ('bazfile.bin', open('bazfile').read(),
                              'image/jpeg'),
                'nonamefile': 'contents of nonamefile field',
            }

        When uploading a file, providing a filename (the first parameter of the
        tuple) is optional but recommended to best mimic behavior of browsers.

        Note that if ``headers`` are supplied, the 'Content-Type' header will
        be overwritten because it depends on the dynamic random boundary string
        which is used to compose the body of the request. The random boundary
        string can be explicitly set with the ``multipart_boundary`` parameter.
        """
        if headers is None:
            headers = self.headers

        extra_kw = {"headers": {}}

        if fields:
            if "body" in urlopen_kw:
                raise TypeError(
                    "request got values for both 'fields' and 'body', can only specify one."
                )

            if encode_multipart:
                body, content_type = encode_multipart_formdata(
                    fields, boundary=multipart_boundary
                )
            else:
                body, content_type = (
                    urlencode(fields),
                    "application/x-www-form-urlencoded",
                )

            extra_kw["body"] = body
            extra_kw["headers"] = {"Content-Type": content_type}

        extra_kw["headers"].update(headers)
        extra_kw.update(urlopen_kw)

        return self.urlopen(method, url, **extra_kw)


if not six.PY2:

    class RequestModule(sys.modules[__name__].__class__):
        def __call__(self, *args, **kwargs):
            """
            If user tries to call this module directly urllib3 v2.x style raise an error to the user
            suggesting they may need urllib3 v2
            """
            raise TypeError(
                "'module' object is not callable\n"
                "urllib3.request() method is not supported in this release, "
                "upgrade to urllib3 v2 to use it\n"
                "see https://urllib3.readthedocs.io/en/stable/v2-migration-guide.html"
            )

    sys.modules[__name__].__class__ = RequestModule

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\actions\action_builder.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

from typing import List, Optional, Union

from selenium.webdriver.remote.command import Command

from . import interaction
from .key_actions import KeyActions
from .key_input import KeyInput
from .pointer_actions import PointerActions
from .pointer_input import PointerInput
from .wheel_actions import WheelActions
from .wheel_input import WheelInput


class ActionBuilder:
    def __init__(
        self,
        driver,
        mouse: Optional[PointerInput] = None,
        wheel: Optional[WheelInput] = None,
        keyboard: Optional[KeyInput] = None,
        duration: int = 250,
    ) -> None:
        mouse = mouse or PointerInput(interaction.POINTER_MOUSE, "mouse")
        keyboard = keyboard or KeyInput(interaction.KEY)
        wheel = wheel or WheelInput(interaction.WHEEL)
        self.devices = [mouse, keyboard, wheel]
        self._key_action = KeyActions(keyboard)
        self._pointer_action = PointerActions(mouse, duration=duration)
        self._wheel_action = WheelActions(wheel)
        self.driver = driver

    def get_device_with(self, name: str) -> Optional[Union["WheelInput", "PointerInput", "KeyInput"]]:
        """Get the device with the given name.

        Parameters:
        -----------
        name : str
            The name of the device to get.

        Returns:
        --------
        Optional[Union[WheelInput, PointerInput, KeyInput]] : The device with the given name.
        """
        return next(filter(lambda x: x == name, self.devices), None)

    @property
    def pointer_inputs(self) -> List[PointerInput]:
        return [device for device in self.devices if device.type == interaction.POINTER]

    @property
    def key_inputs(self) -> List[KeyInput]:
        return [device for device in self.devices if device.type == interaction.KEY]

    @property
    def key_action(self) -> KeyActions:
        return self._key_action

    @property
    def pointer_action(self) -> PointerActions:
        return self._pointer_action

    @property
    def wheel_action(self) -> WheelActions:
        return self._wheel_action

    def add_key_input(self, name: str) -> KeyInput:
        """Add a new key input device to the action builder.

        Parameters:
        -----------
        name : str
            The name of the key input device.

        Returns:
        --------
        KeyInput : The newly created key input device.

        Example:
        --------
        >>> action_builder = ActionBuilder(driver)
        >>> action_builder.add_key_input(name="keyboard2")
        """
        new_input = KeyInput(name)
        self._add_input(new_input)
        return new_input

    def add_pointer_input(self, kind: str, name: str) -> PointerInput:
        """Add a new pointer input device to the action builder.

        Parameters:
        -----------
        kind : str
            The kind of pointer input device.
                - "mouse"
                - "touch"
                - "pen"

        name : str
            The name of the pointer input device.

        Returns:
        --------
        PointerInput : The newly created pointer input device.

        Example:
        --------
        >>> action_builder = ActionBuilder(driver)
        >>> action_builder.add_pointer_input(kind="mouse", name="mouse")
        """
        new_input = PointerInput(kind, name)
        self._add_input(new_input)
        return new_input

    def add_wheel_input(self, name: str) -> WheelInput:
        """Add a new wheel input device to the action builder.

        Parameters:
        -----------
        name : str
            The name of the wheel input device.

        Returns:
        --------
        WheelInput : The newly created wheel input device.

        Example:
        --------
        >>> action_builder = ActionBuilder(driver)
        >>> action_builder.add_wheel_input(name="wheel2")
        """
        new_input = WheelInput(name)
        self._add_input(new_input)
        return new_input

    def perform(self) -> None:
        """Performs all stored actions.

        Example:
        --------
        >>> action_builder = ActionBuilder(driver)
        >>> keyboard = action_builder.key_input
        >>> el = driver.find_element(id: "some_id")
        >>> action_builder.click(el).pause(keyboard).pause(keyboard).pause(keyboard).send_keys("keys").perform()
        """
        enc = {"actions": []}
        for device in self.devices:
            encoded = device.encode()
            if encoded["actions"]:
                enc["actions"].append(encoded)
                device.actions = []
        self.driver.execute(Command.W3C_ACTIONS, enc)

    def clear_actions(self) -> None:
        """Clears actions that are already stored on the remote end.

        Example:
        --------
        >>> action_builder = ActionBuilder(driver)
        >>> keyboard = action_builder.key_input
        >>> el = driver.find_element(By.ID, "some_id")
        >>> action_builder.click(el).pause(keyboard).pause(keyboard).pause(keyboard).send_keys("keys")
        >>> action_builder.clear_actions()
        """
        self.driver.execute(Command.W3C_CLEAR_ACTIONS)

    def _add_input(self, new_input: Union[KeyInput, PointerInput, WheelInput]) -> None:
        """Add a new input device to the action builder.

        Parameters:
        -----------
        new_input : Union[KeyInput, PointerInput, WheelInput]
            The new input device to add.
        """
        self.devices.append(new_input)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\quantum\fermion.py ===
"""Fermionic quantum operators."""

from sympy.core.numbers import Integer
from sympy.core.singleton import S
from sympy.physics.quantum import Operator
from sympy.physics.quantum import HilbertSpace, Ket, Bra
from sympy.functions.special.tensor_functions import KroneckerDelta


__all__ = [
    'FermionOp',
    'FermionFockKet',
    'FermionFockBra'
]


class FermionOp(Operator):
    """A fermionic operator that satisfies {c, Dagger(c)} == 1.

    Parameters
    ==========

    name : str
        A string that labels the fermionic mode.

    annihilation : bool
        A bool that indicates if the fermionic operator is an annihilation
        (True, default value) or creation operator (False)

    Examples
    ========

    >>> from sympy.physics.quantum import Dagger, AntiCommutator
    >>> from sympy.physics.quantum.fermion import FermionOp
    >>> c = FermionOp("c")
    >>> AntiCommutator(c, Dagger(c)).doit()
    1
    """
    @property
    def name(self):
        return self.args[0]

    @property
    def is_annihilation(self):
        return bool(self.args[1])

    @classmethod
    def default_args(self):
        return ("c", True)

    def __new__(cls, *args, **hints):
        if not len(args) in [1, 2]:
            raise ValueError('1 or 2 parameters expected, got %s' % args)

        if len(args) == 1:
            args = (args[0], S.One)

        if len(args) == 2:
            args = (args[0], Integer(args[1]))

        return Operator.__new__(cls, *args)

    def _eval_commutator_FermionOp(self, other, **hints):
        if 'independent' in hints and hints['independent']:
            # [c, d] = 0
            return S.Zero

        return None

    def _eval_anticommutator_FermionOp(self, other, **hints):
        if self.name == other.name:
            # {a^\dagger, a} = 1
            if not self.is_annihilation and other.is_annihilation:
                return S.One

        elif 'independent' in hints and hints['independent']:
            # {c, d} = 2 * c * d, because [c, d] = 0 for independent operators
            return 2 * self * other

        return None

    def _eval_anticommutator_BosonOp(self, other, **hints):
        # because fermions and bosons commute
        return 2 * self * other

    def _eval_commutator_BosonOp(self, other, **hints):
        return S.Zero

    def _eval_adjoint(self):
        return FermionOp(str(self.name), not self.is_annihilation)

    def _print_contents_latex(self, printer, *args):
        if self.is_annihilation:
            return r'{%s}' % str(self.name)
        else:
            return r'{{%s}^\dagger}' % str(self.name)

    def _print_contents(self, printer, *args):
        if self.is_annihilation:
            return r'%s' % str(self.name)
        else:
            return r'Dagger(%s)' % str(self.name)

    def _print_contents_pretty(self, printer, *args):
        from sympy.printing.pretty.stringpict import prettyForm
        pform = printer._print(self.args[0], *args)
        if self.is_annihilation:
            return pform
        else:
            return pform**prettyForm('\N{DAGGER}')

    def _eval_power(self, exp):
        from sympy.core.singleton import S
        if exp == 0:
            return S.One
        elif exp == 1:
            return self
        elif (exp > 1) == True and exp.is_integer == True:
            return S.Zero
        elif (exp < 0) == True or exp.is_integer == False:
            raise ValueError("Fermionic operators can only be raised to a"
                " positive integer power")
        return Operator._eval_power(self, exp)

class FermionFockKet(Ket):
    """Fock state ket for a fermionic mode.

    Parameters
    ==========

    n : Number
        The Fock state number.

    """

    def __new__(cls, n):
        if n not in (0, 1):
            raise ValueError("n must be 0 or 1")
        return Ket.__new__(cls, n)

    @property
    def n(self):
        return self.label[0]

    @classmethod
    def dual_class(self):
        return FermionFockBra

    @classmethod
    def _eval_hilbert_space(cls, label):
        return HilbertSpace()

    def _eval_innerproduct_FermionFockBra(self, bra, **hints):
        return KroneckerDelta(self.n, bra.n)

    def _apply_from_right_to_FermionOp(self, op, **options):
        if op.is_annihilation:
            if self.n == 1:
                return FermionFockKet(0)
            else:
                return S.Zero
        else:
            if self.n == 0:
                return FermionFockKet(1)
            else:
                return S.Zero


class FermionFockBra(Bra):
    """Fock state bra for a fermionic mode.

    Parameters
    ==========

    n : Number
        The Fock state number.

    """

    def __new__(cls, n):
        if n not in (0, 1):
            raise ValueError("n must be 0 or 1")
        return Bra.__new__(cls, n)

    @property
    def n(self):
        return self.label[0]

    @classmethod
    def dual_class(self):
        return FermionFockKet

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\dotenv\cli.py ===
import json
import os
import shlex
import sys
from contextlib import contextmanager
from typing import Any, Dict, IO, Iterator, List, Optional

try:
    import click
except ImportError:
    sys.stderr.write('It seems python-dotenv is not installed with cli option. \n'
                     'Run pip install "python-dotenv[cli]" to fix this.')
    sys.exit(1)

from .main import dotenv_values, set_key, unset_key
from .version import __version__


def enumerate_env() -> Optional[str]:
    """
    Return a path for the ${pwd}/.env file.

    If pwd does not exist, return None.
    """
    try:
        cwd = os.getcwd()
    except FileNotFoundError:
        return None
    path = os.path.join(cwd, '.env')
    return path


@click.group()
@click.option('-f', '--file', default=enumerate_env(),
              type=click.Path(file_okay=True),
              help="Location of the .env file, defaults to .env file in current working directory.")
@click.option('-q', '--quote', default='always',
              type=click.Choice(['always', 'never', 'auto']),
              help="Whether to quote or not the variable values. Default mode is always. This does not affect parsing.")
@click.option('-e', '--export', default=False,
              type=click.BOOL,
              help="Whether to write the dot file as an executable bash script.")
@click.version_option(version=__version__)
@click.pass_context
def cli(ctx: click.Context, file: Any, quote: Any, export: Any) -> None:
    """This script is used to set, get or unset values from a .env file."""
    ctx.obj = {'QUOTE': quote, 'EXPORT': export, 'FILE': file}


@contextmanager
def stream_file(path: os.PathLike) -> Iterator[IO[str]]:
    """
    Open a file and yield the corresponding (decoded) stream.

    Exits with error code 2 if the file cannot be opened.
    """

    try:
        with open(path) as stream:
            yield stream
    except OSError as exc:
        print(f"Error opening env file: {exc}", file=sys.stderr)
        exit(2)


@cli.command()
@click.pass_context
@click.option('--format', default='simple',
              type=click.Choice(['simple', 'json', 'shell', 'export']),
              help="The format in which to display the list. Default format is simple, "
                   "which displays name=value without quotes.")
def list(ctx: click.Context, format: bool) -> None:
    """Display all the stored key/value."""
    file = ctx.obj['FILE']

    with stream_file(file) as stream:
        values = dotenv_values(stream=stream)

    if format == 'json':
        click.echo(json.dumps(values, indent=2, sort_keys=True))
    else:
        prefix = 'export ' if format == 'export' else ''
        for k in sorted(values):
            v = values[k]
            if v is not None:
                if format in ('export', 'shell'):
                    v = shlex.quote(v)
                click.echo(f'{prefix}{k}={v}')


@cli.command()
@click.pass_context
@click.argument('key', required=True)
@click.argument('value', required=True)
def set(ctx: click.Context, key: Any, value: Any) -> None:
    """Store the given key/value."""
    file = ctx.obj['FILE']
    quote = ctx.obj['QUOTE']
    export = ctx.obj['EXPORT']
    success, key, value = set_key(file, key, value, quote, export)
    if success:
        click.echo(f'{key}={value}')
    else:
        exit(1)


@cli.command()
@click.pass_context
@click.argument('key', required=True)
def get(ctx: click.Context, key: Any) -> None:
    """Retrieve the value for the given key."""
    file = ctx.obj['FILE']

    with stream_file(file) as stream:
        values = dotenv_values(stream=stream)

    stored_value = values.get(key)
    if stored_value:
        click.echo(stored_value)
    else:
        exit(1)


@cli.command()
@click.pass_context
@click.argument('key', required=True)
def unset(ctx: click.Context, key: Any) -> None:
    """Removes the given key."""
    file = ctx.obj['FILE']
    quote = ctx.obj['QUOTE']
    success, key = unset_key(file, key, quote)
    if success:
        click.echo(f"Successfully removed {key}")
    else:
        exit(1)


@cli.command(context_settings={'ignore_unknown_options': True})
@click.pass_context
@click.option(
    "--override/--no-override",
    default=True,
    help="Override variables from the environment file with those from the .env file.",
)
@click.argument('commandline', nargs=-1, type=click.UNPROCESSED)
def run(ctx: click.Context, override: bool, commandline: List[str]) -> None:
    """Run command with environment variables present."""
    file = ctx.obj['FILE']
    if not os.path.isfile(file):
        raise click.BadParameter(
            f'Invalid value for \'-f\' "{file}" does not exist.',
            ctx=ctx
        )
    dotenv_as_dict = {
        k: v
        for (k, v) in dotenv_values(file).items()
        if v is not None and (override or k not in os.environ)
    }

    if not commandline:
        click.echo('No command given.')
        exit(1)
    run_command(commandline, dotenv_as_dict)


def run_command(command: List[str], env: Dict[str, str]) -> None:
    """Replace the current process with the specified command.

    Replaces the current process with the specified command and the variables from `env`
    added in the current environment variables.

    Parameters
    ----------
    command: List[str]
        The command and it's parameters
    env: Dict
        The additional environment variables

    Returns
    -------
    None
        This function does not return any value. It replaces the current process with the new one.

    """
    # copy the current environment variables and add the vales from
    # `env`
    cmd_env = os.environ.copy()
    cmd_env.update(env)

    os.execvpe(command[0], args=command, env=cmd_env)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\google\protobuf\message_factory.py ===
# Protocol Buffers - Google's data interchange format
# Copyright 2008 Google Inc.  All rights reserved.
#
# Use of this source code is governed by a BSD-style
# license that can be found in the LICENSE file or at
# https://developers.google.com/open-source/licenses/bsd

"""Provides a factory class for generating dynamic messages.

The easiest way to use this class is if you have access to the FileDescriptor
protos containing the messages you want to create you can just do the following:

message_classes = message_factory.GetMessages(iterable_of_file_descriptors)
my_proto_instance = message_classes['some.proto.package.MessageName']()
"""

__author__ = 'matthewtoia@google.com (Matt Toia)'

import warnings

from google.protobuf import descriptor_pool
from google.protobuf import message
from google.protobuf.internal import api_implementation

if api_implementation.Type() == 'python':
  from google.protobuf.internal import python_message as message_impl
else:
  from google.protobuf.pyext import cpp_message as message_impl  # pylint: disable=g-import-not-at-top


# The type of all Message classes.
_GENERATED_PROTOCOL_MESSAGE_TYPE = message_impl.GeneratedProtocolMessageType


def GetMessageClass(descriptor):
  """Obtains a proto2 message class based on the passed in descriptor.

  Passing a descriptor with a fully qualified name matching a previous
  invocation will cause the same class to be returned.

  Args:
    descriptor: The descriptor to build from.

  Returns:
    A class describing the passed in descriptor.
  """
  concrete_class = getattr(descriptor, '_concrete_class', None)
  if concrete_class:
    return concrete_class
  return _InternalCreateMessageClass(descriptor)


def GetMessageClassesForFiles(files, pool):
  """Gets all the messages from specified files.

  This will find and resolve dependencies, failing if the descriptor
  pool cannot satisfy them.

  This will not return the classes for nested types within those classes, for
  those, use GetMessageClass() on the nested types within their containing
  messages.

  For example, for the message:

  message NestedTypeMessage {
    message NestedType {
      string data = 1;
    }
    NestedType nested = 1;
  }

  NestedTypeMessage will be in the result, but not
  NestedTypeMessage.NestedType.

  Args:
    files: The file names to extract messages from.
    pool: The descriptor pool to find the files including the dependent files.

  Returns:
    A dictionary mapping proto names to the message classes.
  """
  result = {}
  for file_name in files:
    file_desc = pool.FindFileByName(file_name)
    for desc in file_desc.message_types_by_name.values():
      result[desc.full_name] = GetMessageClass(desc)

    # While the extension FieldDescriptors are created by the descriptor pool,
    # the python classes created in the factory need them to be registered
    # explicitly, which is done below.
    #
    # The call to RegisterExtension will specifically check if the
    # extension was already registered on the object and either
    # ignore the registration if the original was the same, or raise
    # an error if they were different.

    for extension in file_desc.extensions_by_name.values():
      _ = GetMessageClass(extension.containing_type)
      if api_implementation.Type() != 'python':
        # TODO: Remove this check here. Duplicate extension
        # register check should be in descriptor_pool.
        if extension is not pool.FindExtensionByNumber(
            extension.containing_type, extension.number
        ):
          raise ValueError('Double registration of Extensions')
      # Recursively load protos for extension field, in order to be able to
      # fully represent the extension. This matches the behavior for regular
      # fields too.
      if extension.message_type:
        GetMessageClass(extension.message_type)
  return result


def _InternalCreateMessageClass(descriptor):
  """Builds a proto2 message class based on the passed in descriptor.

  Args:
    descriptor: The descriptor to build from.

  Returns:
    A class describing the passed in descriptor.
  """
  descriptor_name = descriptor.name
  result_class = _GENERATED_PROTOCOL_MESSAGE_TYPE(
      descriptor_name,
      (message.Message,),
      {
          'DESCRIPTOR': descriptor,
          # If module not set, it wrongly points to message_factory module.
          '__module__': None,
      },
  )
  for field in descriptor.fields:
    if field.message_type:
      GetMessageClass(field.message_type)

  for extension in result_class.DESCRIPTOR.extensions:
    extended_class = GetMessageClass(extension.containing_type)
    if api_implementation.Type() != 'python':
      # TODO: Remove this check here. Duplicate extension
      # register check should be in descriptor_pool.
      pool = extension.containing_type.file.pool
      if extension is not pool.FindExtensionByNumber(
          extension.containing_type, extension.number
      ):
        raise ValueError('Double registration of Extensions')
    if extension.message_type:
      GetMessageClass(extension.message_type)
  return result_class


# Deprecated. Please use GetMessageClass() or GetMessageClassesForFiles()
# method above instead.
class MessageFactory(object):
  """Factory for creating Proto2 messages from descriptors in a pool."""

  def __init__(self, pool=None):
    """Initializes a new factory."""
    self.pool = pool or descriptor_pool.DescriptorPool()


def GetMessages(file_protos, pool=None):
  """Builds a dictionary of all the messages available in a set of files.

  Args:
    file_protos: Iterable of FileDescriptorProto to build messages out of.
    pool: The descriptor pool to add the file protos.

  Returns:
    A dictionary mapping proto names to the message classes. This will include
    any dependent messages as well as any messages defined in the same file as
    a specified message.
  """
  # The cpp implementation of the protocol buffer library requires to add the
  # message in topological order of the dependency graph.
  des_pool = pool or descriptor_pool.DescriptorPool()
  file_by_name = {file_proto.name: file_proto for file_proto in file_protos}

  def _AddFile(file_proto):
    for dependency in file_proto.dependency:
      if dependency in file_by_name:
        # Remove from elements to be visited, in order to cut cycles.
        _AddFile(file_by_name.pop(dependency))
    des_pool.Add(file_proto)

  while file_by_name:
    _AddFile(file_by_name.popitem()[1])
  return GetMessageClassesForFiles(
      [file_proto.name for file_proto in file_protos], des_pool
  )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\workbook\external_link\external.py ===
# Copyright (c) 2010-2024 openpyxl


from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Typed,
    String,
    Bool,
    Integer,
    NoneSet,
    Sequence,
)
from openpyxl.descriptors.excel import Relation
from openpyxl.descriptors.nested import NestedText
from openpyxl.descriptors.sequence import NestedSequence, ValueSequence

from openpyxl.packaging.relationship import (
    Relationship,
    get_rels_path,
    get_dependents
    )
from openpyxl.xml.constants import SHEET_MAIN_NS
from openpyxl.xml.functions import fromstring


"""Manage links to external Workbooks"""


class ExternalCell(Serialisable):

    r = String()
    t = NoneSet(values=(['b', 'd', 'n', 'e', 's', 'str', 'inlineStr']))
    vm = Integer(allow_none=True)
    v = NestedText(allow_none=True, expected_type=str)

    def __init__(self,
                 r=None,
                 t=None,
                 vm=None,
                 v=None,
                ):
        self.r = r
        self.t = t
        self.vm = vm
        self.v = v


class ExternalRow(Serialisable):

    r = Integer()
    cell = Sequence(expected_type=ExternalCell)

    __elements__ = ('cell',)

    def __init__(self,
                 r=(),
                 cell=None,
                ):
        self.r = r
        self.cell = cell


class ExternalSheetData(Serialisable):

    sheetId = Integer()
    refreshError = Bool(allow_none=True)
    row = Sequence(expected_type=ExternalRow)

    __elements__ = ('row',)

    def __init__(self,
                 sheetId=None,
                 refreshError=None,
                 row=(),
                ):
        self.sheetId = sheetId
        self.refreshError = refreshError
        self.row = row


class ExternalSheetDataSet(Serialisable):

    sheetData = Sequence(expected_type=ExternalSheetData, )

    __elements__ = ('sheetData',)

    def __init__(self,
                 sheetData=None,
                ):
        self.sheetData = sheetData


class ExternalSheetNames(Serialisable):

    sheetName = ValueSequence(expected_type=str)

    __elements__ = ('sheetName',)

    def __init__(self,
                 sheetName=(),
                ):
        self.sheetName = sheetName


class ExternalDefinedName(Serialisable):

    tagname = "definedName"

    name = String()
    refersTo = String(allow_none=True)
    sheetId = Integer(allow_none=True)

    def __init__(self,
                 name=None,
                 refersTo=None,
                 sheetId=None,
                ):
        self.name = name
        self.refersTo = refersTo
        self.sheetId = sheetId


class ExternalBook(Serialisable):

    tagname = "externalBook"

    sheetNames = Typed(expected_type=ExternalSheetNames, allow_none=True)
    definedNames = NestedSequence(expected_type=ExternalDefinedName)
    sheetDataSet = Typed(expected_type=ExternalSheetDataSet, allow_none=True)
    id = Relation()

    __elements__ = ('sheetNames', 'definedNames', 'sheetDataSet')

    def __init__(self,
                 sheetNames=None,
                 definedNames=(),
                 sheetDataSet=None,
                 id=None,
                ):
        self.sheetNames = sheetNames
        self.definedNames = definedNames
        self.sheetDataSet = sheetDataSet
        self.id = id


class ExternalLink(Serialisable):

    tagname = "externalLink"

    _id = None
    _path = "/xl/externalLinks/externalLink{0}.xml"
    _rel_type = "externalLink"
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"

    externalBook = Typed(expected_type=ExternalBook, allow_none=True)
    file_link = Typed(expected_type=Relationship, allow_none=True) # link to external file

    __elements__ = ('externalBook', )

    def __init__(self,
                 externalBook=None,
                 ddeLink=None,
                 oleLink=None,
                 extLst=None,
                ):
        self.externalBook = externalBook
        # ignore other items for the moment.


    def to_tree(self):
        node = super().to_tree()
        node.set("xmlns", SHEET_MAIN_NS)
        return node


    @property
    def path(self):
        return self._path.format(self._id)


def read_external_link(archive, book_path):
    src = archive.read(book_path)
    node = fromstring(src)
    book = ExternalLink.from_tree(node)

    link_path = get_rels_path(book_path)
    deps = get_dependents(archive, link_path)
    book.file_link = deps[0]

    return book

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\worksheet\_read_only.py ===
# Copyright (c) 2010-2024 openpyxl

""" Read worksheets on-demand
"""

from .worksheet import Worksheet
from openpyxl.cell.read_only import ReadOnlyCell, EMPTY_CELL
from openpyxl.utils import get_column_letter

from ._reader import WorkSheetParser
from openpyxl.workbook.defined_name import DefinedNameDict


def read_dimension(source):
    parser = WorkSheetParser(source, [])
    return parser.parse_dimensions()


class ReadOnlyWorksheet:

    _min_column = 1
    _min_row = 1
    _max_column = _max_row = None

    # from Standard Worksheet
    # Methods from Worksheet
    cell = Worksheet.cell
    iter_rows = Worksheet.iter_rows
    values = Worksheet.values
    rows = Worksheet.rows
    __getitem__ = Worksheet.__getitem__
    __iter__ = Worksheet.__iter__


    def __init__(self, parent_workbook, title, worksheet_path, shared_strings):
        self.parent = parent_workbook
        self.title = title
        self.sheet_state = 'visible'
        self._current_row = None
        self._worksheet_path = worksheet_path
        self._shared_strings = shared_strings
        self._get_size()
        self.defined_names = DefinedNameDict()


    def _get_size(self):
        src = self._get_source()
        parser = WorkSheetParser(src, [])
        dimensions = parser.parse_dimensions()
        src.close()
        if dimensions is not None:
            self._min_column, self._min_row, self._max_column, self._max_row = dimensions


    def _get_source(self):
        """Parse xml source on demand, must close after use"""
        return self.parent._archive.open(self._worksheet_path)


    def _cells_by_row(self, min_col, min_row, max_col, max_row, values_only=False):
        """
        The source worksheet file may have columns or rows missing.
        Missing cells will be created.
        """
        filler = EMPTY_CELL
        if values_only:
            filler = None

        max_col = max_col or self.max_column
        max_row = max_row or self.max_row
        empty_row = []
        if max_col is not None:
            empty_row = (filler,) * (max_col + 1 - min_col)

        counter = min_row
        idx = 1
        with self._get_source() as src:
            parser = WorkSheetParser(src,
                                     self._shared_strings,
                                     data_only=self.parent.data_only,
                                     epoch=self.parent.epoch,
                                     date_formats=self.parent._date_formats,
                                     timedelta_formats=self.parent._timedelta_formats)

            for idx, row in parser.parse():
                if max_row is not None and idx > max_row:
                    break

                # some rows are missing
                for _ in range(counter, idx):
                    counter += 1
                    yield empty_row

                # return cells from a row
                if counter <= idx:
                    row = self._get_row(row, min_col, max_col, values_only)
                    counter += 1
                    yield row

        if max_row is not None and max_row < idx:
            for _ in range(counter, max_row+1):
                yield empty_row


    def _get_row(self, row, min_col=1, max_col=None, values_only=False):
        """
        Make sure a row contains always the same number of cells or values
        """
        if not row and not max_col: # in case someone wants to force rows where there aren't any
            return ()

        max_col = max_col or  row[-1]['column']
        row_width = max_col + 1 - min_col

        new_row = [EMPTY_CELL] * row_width
        if values_only:
            new_row = [None] * row_width

        for cell in row:
            counter = cell['column']
            if min_col <= counter <= max_col:
                idx = counter - min_col # position in list of cells returned
                new_row[idx] = cell['value']
                if not values_only:
                    new_row[idx] = ReadOnlyCell(self, **cell)

        return tuple(new_row)


    def _get_cell(self, row, column):
        """Cells are returned by a generator which can be empty"""
        for row in self._cells_by_row(column, row, column, row):
            if row:
                return row[0]
        return EMPTY_CELL


    def calculate_dimension(self, force=False):
        if not all([self.max_column, self.max_row]):
            if force:
                self._calculate_dimension()
            else:
                raise ValueError("Worksheet is unsized, use calculate_dimension(force=True)")
        return f"{get_column_letter(self.min_column)}{self.min_row}:{get_column_letter(self.max_column)}{self.max_row}"


    def _calculate_dimension(self):
        """
        Loop through all the cells to get the size of a worksheet.
        Do this only if it is explicitly requested.
        """

        max_col = 0
        for r in self.rows:
            if not r:
                continue
            cell = r[-1]
            max_col = max(max_col, cell.column)

        self._max_row = cell.row
        self._max_column = max_col


    def reset_dimensions(self):
        """
        Remove worksheet dimensions if these are incorrect in the worksheet source.
        NB. This probably indicates a bug in the library or application that created
        the workbook.
        """
        self._max_row = self._max_column = None


    @property
    def min_row(self):
        return self._min_row


    @property
    def max_row(self):
        return self._max_row


    @property
    def min_column(self):
        return self._min_column


    @property
    def max_column(self):
        return self._max_column

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\ntheory\bbp_pi.py ===
'''
This implementation is a heavily modified fixed point implementation of
BBP_formula for calculating the nth position of pi. The original hosted
at: https://web.archive.org/web/20151116045029/http://en.literateprograms.org/Pi_with_the_BBP_formula_(Python)

# Permission is hereby granted, free of charge, to any person obtaining
# a copy of this software and associated documentation files (the
# "Software"), to deal in the Software without restriction, including
# without limitation the rights to use, copy, modify, merge, publish,
# distribute, sub-license, and/or sell copies of the Software, and to
# permit persons to whom the Software is furnished to do so, subject to
# the following conditions:
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
# EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF
# MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY
# CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
# TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE
# SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

Modifications:

1.Once the nth digit and desired number of digits is selected, the
number of digits of working precision is calculated to ensure that
the hexadecimal digits returned are accurate. This is calculated as

    int(math.log(start + prec)/math.log(16) + prec + 3)
    ---------------------------------------   --------
                      /                          /
    number of hex digits           additional digits

This was checked by the following code which completed without
errors (and dig are the digits included in the test_bbp.py file):

    for i in range(0,1000):
     for j in range(1,1000):
      a, b = pi_hex_digits(i, j), dig[i:i+j]
      if a != b:
        print('%s\n%s'%(a,b))

Deceasing the additional digits by 1 generated errors, so '3' is
the smallest additional precision needed to calculate the above
loop without errors. The following trailing 10 digits were also
checked to be accurate (and the times were slightly faster with
some of the constant modifications that were made):

    >> from time import time
    >> t=time();pi_hex_digits(10**2-10 + 1, 10), time()-t
    ('e90c6cc0ac', 0.0)
    >> t=time();pi_hex_digits(10**4-10 + 1, 10), time()-t
    ('26aab49ec6', 0.17100000381469727)
    >> t=time();pi_hex_digits(10**5-10 + 1, 10), time()-t
    ('a22673c1a5', 4.7109999656677246)
    >> t=time();pi_hex_digits(10**6-10 + 1, 10), time()-t
    ('9ffd342362', 59.985999822616577)
    >> t=time();pi_hex_digits(10**7-10 + 1, 10), time()-t
    ('c1a42e06a1', 689.51800012588501)

2. The while loop to evaluate whether the series has converged quits
when the addition amount `dt` has dropped to zero.

3. the formatting string to convert the decimal to hexadecimal is
calculated for the given precision.

4. pi_hex_digits(n) changed to have coefficient to the formula in an
array (perhaps just a matter of preference).

'''

from sympy.utilities.misc import as_int


def _series(j, n, prec=14):

    # Left sum from the bbp algorithm
    s = 0
    D = _dn(n, prec)
    D4 = 4 * D
    d = j
    for k in range(n + 1):
        s += (pow(16, n - k, d) << D4) // d
        d += 8

    # Right sum iterates to infinity for full precision, but we
    # stop at the point where one iteration is beyond the precision
    # specified.

    t = 0
    k = n + 1
    e = D4 - 4 # 4*(D + n - k)
    d = 8 * k + j
    while True:
        dt = (1 << e) // d
        if not dt:
            break
        t += dt
        # k += 1
        e -= 4
        d += 8
    total = s + t

    return total


def pi_hex_digits(n, prec=14):
    """Returns a string containing ``prec`` (default 14) digits
    starting at the nth digit of pi in hex. Counting of digits
    starts at 0 and the decimal is not counted, so for n = 0 the
    returned value starts with 3; n = 1 corresponds to the first
    digit past the decimal point (which in hex is 2).

    Parameters
    ==========

    n : non-negative integer
    prec : non-negative integer. default = 14

    Returns
    =======

    str : Returns a string containing ``prec`` digits
          starting at the nth digit of pi in hex.
          If ``prec`` = 0, returns empty string.

    Raises
    ======

    ValueError
        If ``n`` < 0 or ``prec`` < 0.
        Or ``n`` or ``prec`` is not an integer.

    Examples
    ========

    >>> from sympy.ntheory.bbp_pi import pi_hex_digits
    >>> pi_hex_digits(0)
    '3243f6a8885a30'
    >>> pi_hex_digits(0, 3)
    '324'

    These are consistent with the following results

    >>> import math
    >>> hex(int(math.pi * 2**((14-1)*4)))
    '0x3243f6a8885a30'
    >>> hex(int(math.pi * 2**((3-1)*4)))
    '0x324'

    References
    ==========

    .. [1] http://www.numberworld.org/digits/Pi/
    """
    n, prec = as_int(n), as_int(prec)
    if n < 0:
        raise ValueError('n cannot be negative')
    if prec < 0:
        raise ValueError('prec cannot be negative')
    if prec == 0:
        return ''

    # main of implementation arrays holding formulae coefficients
    n -= 1
    a = [4, 2, 1, 1]
    j = [1, 4, 5, 6]

    #formulae
    D = _dn(n, prec)
    x = + (a[0]*_series(j[0], n, prec)
         - a[1]*_series(j[1], n, prec)
         - a[2]*_series(j[2], n, prec)
         - a[3]*_series(j[3], n, prec)) & (16**D - 1)

    s = ("%0" + "%ix" % prec) % (x // 16**(D - prec))
    return s


def _dn(n, prec):
    # controller for n dependence on precision
    # n = starting digit index
    # prec = the number of total digits to compute
    n += 1  # because we subtract 1 for _series

    # assert int(math.log(n + prec)/math.log(16)) ==\
    #  ((n + prec).bit_length() - 1) // 4
    return ((n + prec).bit_length() - 1) // 4 + prec + 3

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\websocket\_url.py ===
import os
import socket
import struct
from typing import Optional
from urllib.parse import unquote, urlparse
from ._exceptions import WebSocketProxyException

"""
_url.py
websocket - WebSocket client library for Python

Copyright 2024 engn33r

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

__all__ = ["parse_url", "get_proxy_info"]


def parse_url(url: str) -> tuple:
    """
    parse url and the result is tuple of
    (hostname, port, resource path and the flag of secure mode)

    Parameters
    ----------
    url: str
        url string.
    """
    if ":" not in url:
        raise ValueError("url is invalid")

    scheme, url = url.split(":", 1)

    parsed = urlparse(url, scheme="http")
    if parsed.hostname:
        hostname = parsed.hostname
    else:
        raise ValueError("hostname is invalid")
    port = 0
    if parsed.port:
        port = parsed.port

    is_secure = False
    if scheme == "ws":
        if not port:
            port = 80
    elif scheme == "wss":
        is_secure = True
        if not port:
            port = 443
    else:
        raise ValueError("scheme %s is invalid" % scheme)

    if parsed.path:
        resource = parsed.path
    else:
        resource = "/"

    if parsed.query:
        resource += f"?{parsed.query}"

    return hostname, port, resource, is_secure


DEFAULT_NO_PROXY_HOST = ["localhost", "127.0.0.1"]


def _is_ip_address(addr: str) -> bool:
    try:
        socket.inet_aton(addr)
    except socket.error:
        return False
    else:
        return True


def _is_subnet_address(hostname: str) -> bool:
    try:
        addr, netmask = hostname.split("/")
        return _is_ip_address(addr) and 0 <= int(netmask) < 32
    except ValueError:
        return False


def _is_address_in_network(ip: str, net: str) -> bool:
    ipaddr: int = struct.unpack("!I", socket.inet_aton(ip))[0]
    netaddr, netmask = net.split("/")
    netaddr: int = struct.unpack("!I", socket.inet_aton(netaddr))[0]

    netmask = (0xFFFFFFFF << (32 - int(netmask))) & 0xFFFFFFFF
    return ipaddr & netmask == netaddr


def _is_no_proxy_host(hostname: str, no_proxy: Optional[list]) -> bool:
    if not no_proxy:
        if v := os.environ.get("no_proxy", os.environ.get("NO_PROXY", "")).replace(
            " ", ""
        ):
            no_proxy = v.split(",")
    if not no_proxy:
        no_proxy = DEFAULT_NO_PROXY_HOST

    if "*" in no_proxy:
        return True
    if hostname in no_proxy:
        return True
    if _is_ip_address(hostname):
        return any(
            [
                _is_address_in_network(hostname, subnet)
                for subnet in no_proxy
                if _is_subnet_address(subnet)
            ]
        )
    for domain in [domain for domain in no_proxy if domain.startswith(".")]:
        if hostname.endswith(domain):
            return True
    return False


def get_proxy_info(
    hostname: str,
    is_secure: bool,
    proxy_host: Optional[str] = None,
    proxy_port: int = 0,
    proxy_auth: Optional[tuple] = None,
    no_proxy: Optional[list] = None,
    proxy_type: str = "http",
) -> tuple:
    """
    Try to retrieve proxy host and port from environment
    if not provided in options.
    Result is (proxy_host, proxy_port, proxy_auth).
    proxy_auth is tuple of username and password
    of proxy authentication information.

    Parameters
    ----------
    hostname: str
        Websocket server name.
    is_secure: bool
        Is the connection secure? (wss) looks for "https_proxy" in env
        instead of "http_proxy"
    proxy_host: str
        http proxy host name.
    proxy_port: str or int
        http proxy port.
    no_proxy: list
        Whitelisted host names that don't use the proxy.
    proxy_auth: tuple
        HTTP proxy auth information. Tuple of username and password. Default is None.
    proxy_type: str
        Specify the proxy protocol (http, socks4, socks4a, socks5, socks5h). Default is "http".
        Use socks4a or socks5h if you want to send DNS requests through the proxy.
    """
    if _is_no_proxy_host(hostname, no_proxy):
        return None, 0, None

    if proxy_host:
        if not proxy_port:
            raise WebSocketProxyException("Cannot use port 0 when proxy_host specified")
        port = proxy_port
        auth = proxy_auth
        return proxy_host, port, auth

    env_key = "https_proxy" if is_secure else "http_proxy"
    value = os.environ.get(env_key, os.environ.get(env_key.upper(), "")).replace(
        " ", ""
    )
    if value:
        proxy = urlparse(value)
        auth = (
            (unquote(proxy.username), unquote(proxy.password))
            if proxy.username
            else None
        )
        return proxy.hostname, proxy.port, auth

    return None, 0, None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\sam2\prompt_encoder.py ===
# -------------------------------------------------------------------------
# Copyright (R) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
import logging

import torch
from sam2.modeling.sam2_base import SAM2Base
from sam2_utils import compare_tensors_with_tolerance
from torch import nn

logger = logging.getLogger(__name__)


class SAM2PromptEncoder(nn.Module):
    def __init__(self, sam_model: SAM2Base):
        super().__init__()
        self.prompt_encoder = sam_model.sam_prompt_encoder
        self.model = sam_model

    @torch.no_grad()
    def forward(
        self,
        point_coords: torch.Tensor,
        point_labels: torch.Tensor,
        input_masks: torch.Tensor,
        has_input_masks: torch.Tensor,
    ):
        """Encode prompts.

           Args:
            point_coords (torch.Tensor): [L, P, 2] shape and float32 dtype and contains the absolute pixel
                                         coordinate in (x, y) format of the P input points in image of size 1024x1024.
            point_labels (torch.Tensor): shape [L, P] and int32 dtype, where 1 means
                                         positive (foreground), 0 means negative (background), -1 means padding,
                                         2 (box left upper corner), 3 (box right bottom corner).
            input_masks (torch.Tensor): [L, 1, H/4, W/4]. Low resolution mask input to the model.
                                        Typically coming from a previous iteration.
            has_input_masks (torch.Tensor): [L]. 1.0 if input_masks is used, 0.0 otherwise.
        Returns:
            sparse_embeddings (torch.Tensor): [L, P+1, 256], embedding for points and boxes.
            dense_embeddings (torch.Tensor):  [L, 256, 64, 64]. embedding for input masks.
            image_pe (torch.Tensor, optional): [1, 256, 64, 64]. image positional encoding.
        """
        sparse_embeddings = self._embed_points(point_coords, point_labels)
        dense_embeddings = self._embed_masks(input_masks, has_input_masks)
        image_pe = self.prompt_encoder.get_dense_pe()

        return sparse_embeddings, dense_embeddings, image_pe

    def _embed_points(self, point_coords: torch.Tensor, point_labels: torch.Tensor) -> torch.Tensor:
        point_coords = point_coords + 0.5

        padding_point = torch.zeros((point_coords.shape[0], 1, 2), device=point_coords.device)
        padding_label = -torch.ones((point_labels.shape[0], 1), device=point_labels.device)
        point_coords = torch.cat([point_coords, padding_point], dim=1)
        point_labels = torch.cat([point_labels, padding_label], dim=1)

        # Note that the input coordinates are based on image size 1024x1024. Here we normalize it to [0.0, 1.0).
        point_coords[:, :, 0] = point_coords[:, :, 0] / self.model.image_size
        point_coords[:, :, 1] = point_coords[:, :, 1] / self.model.image_size

        point_embedding = self.prompt_encoder.pe_layer._pe_encoding(point_coords)
        point_labels = point_labels.unsqueeze(-1).expand_as(point_embedding)

        point_embedding = point_embedding * (point_labels != -1)
        point_embedding = point_embedding + self.prompt_encoder.not_a_point_embed.weight * (point_labels == -1)

        for i in range(self.prompt_encoder.num_point_embeddings):
            point_embedding = point_embedding + self.prompt_encoder.point_embeddings[i].weight * (point_labels == i)

        return point_embedding

    def _embed_masks(self, input_masks: torch.Tensor, has_input_masks: torch.Tensor) -> torch.Tensor:
        mask_embedding = self.prompt_encoder.mask_downscaling(input_masks)
        no_mask_embedding = self.prompt_encoder.no_mask_embed.weight.reshape(1, -1, 1, 1)
        logger.info("no_mask_embedding.shape: %s", no_mask_embedding.shape)
        mask_embedding = has_input_masks * mask_embedding + (1.0 - has_input_masks) * no_mask_embedding
        logger.info("mask_embedding.shape: %s", mask_embedding.shape)
        return mask_embedding


def export_prompt_encoder_onnx(
    sam2_model: SAM2Base,
    onnx_model_path: str,
):
    sam2_prompt_encoder = SAM2PromptEncoder(sam2_model).cpu()

    num_labels = 2
    num_points = 3
    point_coords = torch.randint(low=0, high=1024, size=(num_labels, num_points, 2), dtype=torch.float)
    point_labels = torch.randint(low=0, high=1, size=(num_labels, num_points), dtype=torch.int32)
    input_masks = torch.zeros(num_labels, 1, 256, 256, dtype=torch.float)
    has_input_masks = torch.ones(1, dtype=torch.float)

    sparse_embeddings, dense_embeddings, image_pe = sam2_prompt_encoder(
        point_coords, point_labels, input_masks, has_input_masks
    )

    logger.info("point_coords.shape: %s", point_coords.shape)
    logger.info("point_labels.shape: %s", point_labels.shape)
    logger.info("input_masks.shape: %s", input_masks.shape)
    logger.info("has_input_masks.shape: %s", has_input_masks.shape)

    logger.info("sparse_embeddings.shape: %s", sparse_embeddings.shape)
    logger.info("dense_embeddings.shape: %s", dense_embeddings.shape)
    logger.info("image_pe.shape: %s", image_pe.shape)

    torch.onnx.export(
        sam2_prompt_encoder,
        (point_coords, point_labels, input_masks, has_input_masks),
        onnx_model_path,
        export_params=True,
        opset_version=18,
        do_constant_folding=True,
        input_names=["point_coords", "point_labels", "input_masks", "has_input_masks"],
        output_names=["sparse_embeddings", "dense_embeddings", "image_pe"],
        dynamic_axes={
            "point_coords": {0: "num_labels", 1: "num_points"},
            "point_labels": {0: "num_labels", 1: "num_points"},
            "input_masks": {0: "num_labels"},
            "sparse_embeddings": {0: "num_labels", 1: "num_points+1"},
            "dense_embeddings": {0: "num_labels"},
        },
    )

    print("prompt encoder onnx model saved to ", onnx_model_path)


def test_prompt_encoder_onnx(
    sam2_model: SAM2Base,
    onnx_model_path: str,
):
    sam2_prompt_encoder = SAM2PromptEncoder(sam2_model).cpu()

    num_labels = 1
    num_points = 5
    point_coords = torch.randint(low=0, high=1024, size=(num_labels, num_points, 2), dtype=torch.float)
    point_labels = torch.randint(low=0, high=1, size=(num_labels, num_points), dtype=torch.int32)
    input_masks = torch.rand(num_labels, 1, 256, 256, dtype=torch.float)
    has_input_masks = torch.ones(1, dtype=torch.float)

    sparse_embeddings, dense_embeddings, image_pe = sam2_prompt_encoder(
        point_coords, point_labels, input_masks, has_input_masks
    )

    import onnxruntime

    ort_session = onnxruntime.InferenceSession(onnx_model_path, providers=["CPUExecutionProvider"])

    model_inputs = ort_session.get_inputs()
    input_names = [model_inputs[i].name for i in range(len(model_inputs))]
    logger.info("input_names: %s", input_names)

    model_outputs = ort_session.get_outputs()
    output_names = [model_outputs[i].name for i in range(len(model_outputs))]
    logger.info("output_names: %s", output_names)

    outputs = ort_session.run(
        output_names,
        {
            "point_coords": point_coords.numpy(),
            "point_labels": point_labels.numpy(),
            "input_masks": input_masks.numpy(),
            "has_input_masks": has_input_masks.numpy(),
        },
    )

    for i, output_name in enumerate(output_names):
        logger.info("output %s shape: %s", output_name, outputs[i].shape)

    ort_sparse_embeddings, ort_dense_embeddings, ort_image_pe = outputs
    if (
        compare_tensors_with_tolerance(
            "sparse_embeddings",
            sparse_embeddings,
            torch.tensor(ort_sparse_embeddings),
            mismatch_percentage_tolerance=0.2,
        )
        and compare_tensors_with_tolerance(
            "dense_embeddings", dense_embeddings, torch.tensor(ort_dense_embeddings), mismatch_percentage_tolerance=0.2
        )
        and compare_tensors_with_tolerance(
            "image_pe", image_pe, torch.tensor(ort_image_pe), mismatch_percentage_tolerance=0.2
        )
    ):
        print(f"onnx model has been verified: {onnx_model_path}")
    else:
        print(f"onnx model verification failed: {onnx_model_path}")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\workbook\defined_name.py ===
# Copyright (c) 2010-2024 openpyxl

from collections import defaultdict
import re

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Alias,
    String,
    Integer,
    Bool,
    Sequence,
    Descriptor,
)
from openpyxl.compat import safe_string
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import SHEETRANGE_RE

RESERVED = frozenset(["Print_Area", "Print_Titles", "Criteria",
                      "_FilterDatabase", "Extract", "Consolidate_Area",
                      "Sheet_Title"])

_names = "|".join(RESERVED)
RESERVED_REGEX = re.compile(r"^_xlnm\.(?P<name>{0})".format(_names))


class DefinedName(Serialisable):

    tagname = "definedName"

    name = String() # unique per workbook/worksheet
    comment = String(allow_none=True)
    customMenu = String(allow_none=True)
    description = String(allow_none=True)
    help = String(allow_none=True)
    statusBar = String(allow_none=True)
    localSheetId = Integer(allow_none=True)
    hidden = Bool(allow_none=True)
    function = Bool(allow_none=True)
    vbProcedure = Bool(allow_none=True)
    xlm = Bool(allow_none=True)
    functionGroupId = Integer(allow_none=True)
    shortcutKey = String(allow_none=True)
    publishToServer = Bool(allow_none=True)
    workbookParameter = Bool(allow_none=True)
    attr_text = Descriptor()
    value = Alias("attr_text")


    def __init__(self,
                 name=None,
                 comment=None,
                 customMenu=None,
                 description=None,
                 help=None,
                 statusBar=None,
                 localSheetId=None,
                 hidden=None,
                 function=None,
                 vbProcedure=None,
                 xlm=None,
                 functionGroupId=None,
                 shortcutKey=None,
                 publishToServer=None,
                 workbookParameter=None,
                 attr_text=None
                ):
        self.name = name
        self.comment = comment
        self.customMenu = customMenu
        self.description = description
        self.help = help
        self.statusBar = statusBar
        self.localSheetId = localSheetId
        self.hidden = hidden
        self.function = function
        self.vbProcedure = vbProcedure
        self.xlm = xlm
        self.functionGroupId = functionGroupId
        self.shortcutKey = shortcutKey
        self.publishToServer = publishToServer
        self.workbookParameter = workbookParameter
        self.attr_text = attr_text


    @property
    def type(self):
        tok = Tokenizer("=" + self.value)
        parsed = tok.items[0]
        if parsed.type == "OPERAND":
            return parsed.subtype
        return parsed.type


    @property
    def destinations(self):
        if self.type == "RANGE":
            tok = Tokenizer("=" + self.value)
            for part in tok.items:
                if part.subtype == "RANGE":
                    m = SHEETRANGE_RE.match(part.value)
                    sheetname = m.group('notquoted') or m.group('quoted')
                    yield sheetname, m.group('cells')


    @property
    def is_reserved(self):
        m = RESERVED_REGEX.match(self.name)
        if m:
            return m.group("name")


    @property
    def is_external(self):
        return re.compile(r"^\[\d+\].*").match(self.value) is not None


    def __iter__(self):
        for key in self.__attrs__:
            if key == "attr_text":
                continue
            v = getattr(self, key)
            if v is not None:
                if v in RESERVED:
                    v = "_xlnm." + v
                yield key, safe_string(v)


class DefinedNameDict(dict):

    """
    Utility class for storing defined names.
    Allows access by name and separation of global and scoped names
    """

    def __setitem__(self, key, value):
        if not isinstance(value, DefinedName):
            raise TypeError("Value must be a an instance of DefinedName")
        elif value.name != key:
            raise ValueError("Key must be the same as the name")
        super().__setitem__(key, value)


    def add(self, value):
        """
        Add names without worrying about key and name matching.
        """
        self[value.name] = value


class DefinedNameList(Serialisable):

    tagname = "definedNames"

    definedName = Sequence(expected_type=DefinedName)


    def __init__(self, definedName=()):
        self.definedName = definedName


    def by_sheet(self):
        """
        Break names down into sheet locals and globals
        """
        names = defaultdict(DefinedNameDict)
        for defn in self.definedName:
            if defn.localSheetId is None:
                if defn.name in ("_xlnm.Print_Titles", "_xlnm.Print_Area", "_xlnm._FilterDatabase"):
                    continue
                names["global"][defn.name] = defn
            else:
                sheet = int(defn.localSheetId)
                names[sheet][defn.name] = defn
        return names


    def _duplicate(self, defn):
        """
        Check for whether DefinedName with the same name and scope already
        exists
        """
        for d in self.definedName:
            if d.name == defn.name and d.localSheetId == defn.localSheetId:
                return True


    def __len__(self):
        return len(self.definedName)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\ops\mask_ops.py ===
"""
Ops for masked arrays.
"""
from __future__ import annotations

import numpy as np

from pandas._libs import (
    lib,
    missing as libmissing,
)


def kleene_or(
    left: bool | np.ndarray | libmissing.NAType,
    right: bool | np.ndarray | libmissing.NAType,
    left_mask: np.ndarray | None,
    right_mask: np.ndarray | None,
):
    """
    Boolean ``or`` using Kleene logic.

    Values are NA where we have ``NA | NA`` or ``NA | False``.
    ``NA | True`` is considered True.

    Parameters
    ----------
    left, right : ndarray, NA, or bool
        The values of the array.
    left_mask, right_mask : ndarray, optional
        The masks. Only one of these may be None, which implies that
        the associated `left` or `right` value is a scalar.

    Returns
    -------
    result, mask: ndarray[bool]
        The result of the logical or, and the new mask.
    """
    # To reduce the number of cases, we ensure that `left` & `left_mask`
    # always come from an array, not a scalar. This is safe, since
    # A | B == B | A
    if left_mask is None:
        return kleene_or(right, left, right_mask, left_mask)

    if not isinstance(left, np.ndarray):
        raise TypeError("Either `left` or `right` need to be a np.ndarray.")

    raise_for_nan(right, method="or")

    if right is libmissing.NA:
        result = left.copy()
    else:
        result = left | right

    if right_mask is not None:
        # output is unknown where (False & NA), (NA & False), (NA & NA)
        left_false = ~(left | left_mask)
        right_false = ~(right | right_mask)
        mask = (
            (left_false & right_mask)
            | (right_false & left_mask)
            | (left_mask & right_mask)
        )
    else:
        if right is True:
            mask = np.zeros_like(left_mask)
        elif right is libmissing.NA:
            mask = (~left & ~left_mask) | left_mask
        else:
            # False
            mask = left_mask.copy()

    return result, mask


def kleene_xor(
    left: bool | np.ndarray | libmissing.NAType,
    right: bool | np.ndarray | libmissing.NAType,
    left_mask: np.ndarray | None,
    right_mask: np.ndarray | None,
):
    """
    Boolean ``xor`` using Kleene logic.

    This is the same as ``or``, with the following adjustments

    * True, True -> False
    * True, NA   -> NA

    Parameters
    ----------
    left, right : ndarray, NA, or bool
        The values of the array.
    left_mask, right_mask : ndarray, optional
        The masks. Only one of these may be None, which implies that
        the associated `left` or `right` value is a scalar.

    Returns
    -------
    result, mask: ndarray[bool]
        The result of the logical xor, and the new mask.
    """
    # To reduce the number of cases, we ensure that `left` & `left_mask`
    # always come from an array, not a scalar. This is safe, since
    # A ^ B == B ^ A
    if left_mask is None:
        return kleene_xor(right, left, right_mask, left_mask)

    if not isinstance(left, np.ndarray):
        raise TypeError("Either `left` or `right` need to be a np.ndarray.")

    raise_for_nan(right, method="xor")
    if right is libmissing.NA:
        result = np.zeros_like(left)
    else:
        result = left ^ right

    if right_mask is None:
        if right is libmissing.NA:
            mask = np.ones_like(left_mask)
        else:
            mask = left_mask.copy()
    else:
        mask = left_mask | right_mask

    return result, mask


def kleene_and(
    left: bool | libmissing.NAType | np.ndarray,
    right: bool | libmissing.NAType | np.ndarray,
    left_mask: np.ndarray | None,
    right_mask: np.ndarray | None,
):
    """
    Boolean ``and`` using Kleene logic.

    Values are ``NA`` for ``NA & NA`` or ``True & NA``.

    Parameters
    ----------
    left, right : ndarray, NA, or bool
        The values of the array.
    left_mask, right_mask : ndarray, optional
        The masks. Only one of these may be None, which implies that
        the associated `left` or `right` value is a scalar.

    Returns
    -------
    result, mask: ndarray[bool]
        The result of the logical xor, and the new mask.
    """
    # To reduce the number of cases, we ensure that `left` & `left_mask`
    # always come from an array, not a scalar. This is safe, since
    # A & B == B & A
    if left_mask is None:
        return kleene_and(right, left, right_mask, left_mask)

    if not isinstance(left, np.ndarray):
        raise TypeError("Either `left` or `right` need to be a np.ndarray.")
    raise_for_nan(right, method="and")

    if right is libmissing.NA:
        result = np.zeros_like(left)
    else:
        result = left & right

    if right_mask is None:
        # Scalar `right`
        if right is libmissing.NA:
            mask = (left & ~left_mask) | left_mask

        else:
            mask = left_mask.copy()
            if right is False:
                # unmask everything
                mask[:] = False
    else:
        # unmask where either left or right is False
        left_false = ~(left | left_mask)
        right_false = ~(right | right_mask)
        mask = (left_mask & ~right_false) | (right_mask & ~left_false)

    return result, mask


def raise_for_nan(value, method: str) -> None:
    if lib.is_float(value) and np.isnan(value):
        raise ValueError(f"Cannot perform logical '{method}' with floating NaN")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\testing\suite\test_unicode_ddl.py ===
# testing/suite/test_unicode_ddl.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors


from sqlalchemy import desc
from sqlalchemy import ForeignKey
from sqlalchemy import Integer
from sqlalchemy import MetaData
from sqlalchemy import testing
from sqlalchemy.testing import eq_
from sqlalchemy.testing import fixtures
from sqlalchemy.testing.schema import Column
from sqlalchemy.testing.schema import Table


class UnicodeSchemaTest(fixtures.TablesTest):
    __requires__ = ("unicode_ddl",)
    __backend__ = True

    @classmethod
    def define_tables(cls, metadata):
        global t1, t2, t3

        t1 = Table(
            "unitable1",
            metadata,
            Column("méil", Integer, primary_key=True),
            Column("\u6e2c\u8a66", Integer),
            test_needs_fk=True,
        )
        t2 = Table(
            "Unitéble2",
            metadata,
            Column("méil", Integer, primary_key=True, key="a"),
            Column(
                "\u6e2c\u8a66",
                Integer,
                ForeignKey("unitable1.méil"),
                key="b",
            ),
            test_needs_fk=True,
        )

        # Few DBs support Unicode foreign keys
        if testing.against("sqlite"):
            t3 = Table(
                "\u6e2c\u8a66",
                metadata,
                Column(
                    "\u6e2c\u8a66_id",
                    Integer,
                    primary_key=True,
                    autoincrement=False,
                ),
                Column(
                    "unitable1_\u6e2c\u8a66",
                    Integer,
                    ForeignKey("unitable1.\u6e2c\u8a66"),
                ),
                Column("Unitéble2_b", Integer, ForeignKey("Unitéble2.b")),
                Column(
                    "\u6e2c\u8a66_self",
                    Integer,
                    ForeignKey("\u6e2c\u8a66.\u6e2c\u8a66_id"),
                ),
                test_needs_fk=True,
            )
        else:
            t3 = Table(
                "\u6e2c\u8a66",
                metadata,
                Column(
                    "\u6e2c\u8a66_id",
                    Integer,
                    primary_key=True,
                    autoincrement=False,
                ),
                Column("unitable1_\u6e2c\u8a66", Integer),
                Column("Unitéble2_b", Integer),
                Column("\u6e2c\u8a66_self", Integer),
                test_needs_fk=True,
            )

    def test_insert(self, connection):
        connection.execute(t1.insert(), {"méil": 1, "\u6e2c\u8a66": 5})
        connection.execute(t2.insert(), {"a": 1, "b": 1})
        connection.execute(
            t3.insert(),
            {
                "\u6e2c\u8a66_id": 1,
                "unitable1_\u6e2c\u8a66": 5,
                "Unitéble2_b": 1,
                "\u6e2c\u8a66_self": 1,
            },
        )

        eq_(connection.execute(t1.select()).fetchall(), [(1, 5)])
        eq_(connection.execute(t2.select()).fetchall(), [(1, 1)])
        eq_(connection.execute(t3.select()).fetchall(), [(1, 5, 1, 1)])

    def test_col_targeting(self, connection):
        connection.execute(t1.insert(), {"méil": 1, "\u6e2c\u8a66": 5})
        connection.execute(t2.insert(), {"a": 1, "b": 1})
        connection.execute(
            t3.insert(),
            {
                "\u6e2c\u8a66_id": 1,
                "unitable1_\u6e2c\u8a66": 5,
                "Unitéble2_b": 1,
                "\u6e2c\u8a66_self": 1,
            },
        )

        row = connection.execute(t1.select()).first()
        eq_(row._mapping[t1.c["méil"]], 1)
        eq_(row._mapping[t1.c["\u6e2c\u8a66"]], 5)

        row = connection.execute(t2.select()).first()
        eq_(row._mapping[t2.c["a"]], 1)
        eq_(row._mapping[t2.c["b"]], 1)

        row = connection.execute(t3.select()).first()
        eq_(row._mapping[t3.c["\u6e2c\u8a66_id"]], 1)
        eq_(row._mapping[t3.c["unitable1_\u6e2c\u8a66"]], 5)
        eq_(row._mapping[t3.c["Unitéble2_b"]], 1)
        eq_(row._mapping[t3.c["\u6e2c\u8a66_self"]], 1)

    def test_reflect(self, connection):
        connection.execute(t1.insert(), {"méil": 2, "\u6e2c\u8a66": 7})
        connection.execute(t2.insert(), {"a": 2, "b": 2})
        connection.execute(
            t3.insert(),
            {
                "\u6e2c\u8a66_id": 2,
                "unitable1_\u6e2c\u8a66": 7,
                "Unitéble2_b": 2,
                "\u6e2c\u8a66_self": 2,
            },
        )

        meta = MetaData()
        tt1 = Table(t1.name, meta, autoload_with=connection)
        tt2 = Table(t2.name, meta, autoload_with=connection)
        tt3 = Table(t3.name, meta, autoload_with=connection)

        connection.execute(tt1.insert(), {"méil": 1, "\u6e2c\u8a66": 5})
        connection.execute(tt2.insert(), {"méil": 1, "\u6e2c\u8a66": 1})
        connection.execute(
            tt3.insert(),
            {
                "\u6e2c\u8a66_id": 1,
                "unitable1_\u6e2c\u8a66": 5,
                "Unitéble2_b": 1,
                "\u6e2c\u8a66_self": 1,
            },
        )

        eq_(
            connection.execute(tt1.select().order_by(desc("méil"))).fetchall(),
            [(2, 7), (1, 5)],
        )
        eq_(
            connection.execute(tt2.select().order_by(desc("méil"))).fetchall(),
            [(2, 2), (1, 1)],
        )
        eq_(
            connection.execute(
                tt3.select().order_by(desc("\u6e2c\u8a66_id"))
            ).fetchall(),
            [(2, 7, 2, 2), (1, 5, 1, 1)],
        )

    def test_repr(self):
        meta = MetaData()
        t = Table("\u6e2c\u8a66", meta, Column("\u6e2c\u8a66_id", Integer))
        eq_(
            repr(t),
            (
                "Table('測試', MetaData(), "
                "Column('測試_id', Integer(), "
                "table=<測試>), "
                "schema=None)"
            ),
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\vector\parametricregion.py ===
from functools import singledispatch
from sympy.core.numbers import pi
from sympy.functions.elementary.trigonometric import tan
from sympy.simplify import trigsimp
from sympy.core import Basic, Tuple
from sympy.core.symbol import _symbol
from sympy.solvers import solve
from sympy.geometry import Point, Segment, Curve, Ellipse, Polygon
from sympy.vector import ImplicitRegion


class ParametricRegion(Basic):
    """
    Represents a parametric region in space.

    Examples
    ========

    >>> from sympy import cos, sin, pi
    >>> from sympy.abc import r, theta, t, a, b, x, y
    >>> from sympy.vector import ParametricRegion

    >>> ParametricRegion((t, t**2), (t, -1, 2))
    ParametricRegion((t, t**2), (t, -1, 2))
    >>> ParametricRegion((x, y), (x, 3, 4), (y, 5, 6))
    ParametricRegion((x, y), (x, 3, 4), (y, 5, 6))
    >>> ParametricRegion((r*cos(theta), r*sin(theta)), (r, -2, 2), (theta, 0, pi))
    ParametricRegion((r*cos(theta), r*sin(theta)), (r, -2, 2), (theta, 0, pi))
    >>> ParametricRegion((a*cos(t), b*sin(t)), t)
    ParametricRegion((a*cos(t), b*sin(t)), t)

    >>> circle = ParametricRegion((r*cos(theta), r*sin(theta)), r, (theta, 0, pi))
    >>> circle.parameters
    (r, theta)
    >>> circle.definition
    (r*cos(theta), r*sin(theta))
    >>> circle.limits
    {theta: (0, pi)}

    Dimension of a parametric region determines whether a region is a curve, surface
    or volume region. It does not represent its dimensions in space.

    >>> circle.dimensions
    1

    Parameters
    ==========

    definition : tuple to define base scalars in terms of parameters.

    bounds : Parameter or a tuple of length 3 to define parameter and corresponding lower and upper bound.

    """
    def __new__(cls, definition, *bounds):
        parameters = ()
        limits = {}

        if not isinstance(bounds, Tuple):
            bounds = Tuple(*bounds)

        for bound in bounds:
            if isinstance(bound, (tuple, Tuple)):
                if len(bound) != 3:
                    raise ValueError("Tuple should be in the form (parameter, lowerbound, upperbound)")
                parameters += (bound[0],)
                limits[bound[0]] = (bound[1], bound[2])
            else:
                parameters += (bound,)

        if not isinstance(definition, (tuple, Tuple)):
            definition = (definition,)

        obj = super().__new__(cls, Tuple(*definition), *bounds)
        obj._parameters = parameters
        obj._limits = limits

        return obj

    @property
    def definition(self):
        return self.args[0]

    @property
    def limits(self):
        return self._limits

    @property
    def parameters(self):
        return self._parameters

    @property
    def dimensions(self):
        return len(self.limits)


@singledispatch
def parametric_region_list(reg):
    """
    Returns a list of ParametricRegion objects representing the geometric region.

    Examples
    ========

    >>> from sympy.abc import t
    >>> from sympy.vector import parametric_region_list
    >>> from sympy.geometry import Point, Curve, Ellipse, Segment, Polygon

    >>> p = Point(2, 5)
    >>> parametric_region_list(p)
    [ParametricRegion((2, 5))]

    >>> c = Curve((t**3, 4*t), (t, -3, 4))
    >>> parametric_region_list(c)
    [ParametricRegion((t**3, 4*t), (t, -3, 4))]

    >>> e = Ellipse(Point(1, 3), 2, 3)
    >>> parametric_region_list(e)
    [ParametricRegion((2*cos(t) + 1, 3*sin(t) + 3), (t, 0, 2*pi))]

    >>> s = Segment(Point(1, 3), Point(2, 6))
    >>> parametric_region_list(s)
    [ParametricRegion((t + 1, 3*t + 3), (t, 0, 1))]

    >>> p1, p2, p3, p4 = [(0, 1), (2, -3), (5, 3), (-2, 3)]
    >>> poly = Polygon(p1, p2, p3, p4)
    >>> parametric_region_list(poly)
    [ParametricRegion((2*t, 1 - 4*t), (t, 0, 1)), ParametricRegion((3*t + 2, 6*t - 3), (t, 0, 1)),\
     ParametricRegion((5 - 7*t, 3), (t, 0, 1)), ParametricRegion((2*t - 2, 3 - 2*t),  (t, 0, 1))]

    """
    raise ValueError("SymPy cannot determine parametric representation of the region.")


@parametric_region_list.register(Point)
def _(obj):
    return [ParametricRegion(obj.args)]


@parametric_region_list.register(Curve)  # type: ignore
def _(obj):
    definition = obj.arbitrary_point(obj.parameter).args
    bounds = obj.limits
    return [ParametricRegion(definition, bounds)]


@parametric_region_list.register(Ellipse) # type: ignore
def _(obj, parameter='t'):
    definition = obj.arbitrary_point(parameter).args
    t = _symbol(parameter, real=True)
    bounds = (t, 0, 2*pi)
    return [ParametricRegion(definition, bounds)]


@parametric_region_list.register(Segment) # type: ignore
def _(obj, parameter='t'):
    t = _symbol(parameter, real=True)
    definition = obj.arbitrary_point(t).args

    for i in range(0, 3):
        lower_bound = solve(definition[i] - obj.points[0].args[i], t)
        upper_bound = solve(definition[i] - obj.points[1].args[i], t)

        if len(lower_bound) == 1 and len(upper_bound) == 1:
            bounds = t, lower_bound[0], upper_bound[0]
            break

    definition_tuple = obj.arbitrary_point(parameter).args
    return [ParametricRegion(definition_tuple, bounds)]


@parametric_region_list.register(Polygon) # type: ignore
def _(obj, parameter='t'):
    l = [parametric_region_list(side, parameter)[0] for side in obj.sides]
    return l


@parametric_region_list.register(ImplicitRegion) # type: ignore
def _(obj, parameters=('t', 's')):
    definition = obj.rational_parametrization(parameters)
    bounds = []

    for i in range(len(obj.variables) - 1):
        # Each parameter is replaced by its tangent to simplify integration
        parameter = _symbol(parameters[i], real=True)
        definition = [trigsimp(elem.subs(parameter, tan(parameter/2))) for elem in definition]
        bounds.append((parameter, 0, 2*pi),)

    definition = Tuple(*definition)
    return [ParametricRegion(definition, *bounds)]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\test_highlevel_serve_listeners.py ===
from __future__ import annotations

import errno
from functools import partial
from typing import TYPE_CHECKING, NoReturn, cast

import attrs

import trio
from trio import Nursery, StapledStream, TaskStatus
from trio.testing import (
    Matcher,
    MemoryReceiveStream,
    MemorySendStream,
    MockClock,
    RaisesGroup,
    memory_stream_pair,
    wait_all_tasks_blocked,
)

if TYPE_CHECKING:
    from collections.abc import Awaitable, Callable

    import pytest

    from trio._channel import MemoryReceiveChannel, MemorySendChannel
    from trio.abc import Stream

# types are somewhat tentative - I just bruteforced them until I got something that didn't
# give errors
StapledMemoryStream = StapledStream[MemorySendStream, MemoryReceiveStream]


@attrs.define(eq=False, slots=False)
class MemoryListener(trio.abc.Listener[StapledMemoryStream]):
    closed: bool = False
    accepted_streams: list[trio.abc.Stream] = attrs.Factory(list)
    queued_streams: tuple[
        MemorySendChannel[StapledMemoryStream],
        MemoryReceiveChannel[StapledMemoryStream],
    ] = attrs.Factory(lambda: trio.open_memory_channel[StapledMemoryStream](1))
    accept_hook: Callable[[], Awaitable[object]] | None = None

    async def connect(self) -> StapledMemoryStream:
        assert not self.closed
        client, server = memory_stream_pair()
        await self.queued_streams[0].send(server)
        return client

    async def accept(self) -> StapledMemoryStream:
        await trio.lowlevel.checkpoint()
        assert not self.closed
        if self.accept_hook is not None:
            await self.accept_hook()
        stream = await self.queued_streams[1].receive()
        self.accepted_streams.append(stream)
        return stream

    async def aclose(self) -> None:
        self.closed = True
        await trio.lowlevel.checkpoint()


async def test_serve_listeners_basic() -> None:
    listeners = [MemoryListener(), MemoryListener()]

    record = []

    def close_hook() -> None:
        # Make sure this is a forceful close
        assert trio.current_effective_deadline() == float("-inf")
        record.append("closed")

    async def handler(stream: StapledMemoryStream) -> None:
        await stream.send_all(b"123")
        assert await stream.receive_some(10) == b"456"
        stream.send_stream.close_hook = close_hook
        stream.receive_stream.close_hook = close_hook

    async def client(listener: MemoryListener) -> None:
        s = await listener.connect()
        assert await s.receive_some(10) == b"123"
        await s.send_all(b"456")

    async def do_tests(parent_nursery: Nursery) -> None:
        async with trio.open_nursery() as nursery:
            for listener in listeners:
                for _ in range(3):
                    nursery.start_soon(client, listener)

        await wait_all_tasks_blocked()

        # verifies that all 6 streams x 2 directions each were closed ok
        assert len(record) == 12

        parent_nursery.cancel_scope.cancel()

    async with trio.open_nursery() as nursery:
        value = await nursery.start(
            trio.serve_listeners,
            handler,
            listeners,
        )
        assert isinstance(value, list)
        l2 = cast("list[MemoryListener]", value)
        assert l2 == listeners
        # This is just split into another function because gh-136 isn't
        # implemented yet
        nursery.start_soon(do_tests, nursery)

    for listener in listeners:
        assert listener.closed


async def test_serve_listeners_accept_unrecognized_error() -> None:
    for error in [KeyError(), OSError(errno.ECONNABORTED, "ECONNABORTED")]:
        listener = MemoryListener()

        async def raise_error() -> NoReturn:
            raise error  # noqa: B023  # Set from loop

        def check_error(e: BaseException) -> bool:
            return e is error  # noqa: B023

        listener.accept_hook = raise_error

        with RaisesGroup(Matcher(check=check_error)):
            await trio.serve_listeners(None, [listener])  # type: ignore[arg-type]


async def test_serve_listeners_accept_capacity_error(
    autojump_clock: MockClock,
    caplog: pytest.LogCaptureFixture,
) -> None:
    listener = MemoryListener()

    async def raise_EMFILE() -> NoReturn:
        raise OSError(errno.EMFILE, "out of file descriptors")

    listener.accept_hook = raise_EMFILE

    # It retries every 100 ms, so in 950 ms it will retry at 0, 100, ..., 900
    # = 10 times total
    with trio.move_on_after(0.950):
        await trio.serve_listeners(None, [listener])  # type: ignore[arg-type]

    assert len(caplog.records) == 10
    for record in caplog.records:
        assert "retrying" in record.msg
        assert record.exc_info is not None
        assert isinstance(record.exc_info[1], OSError)
        assert record.exc_info[1].errno == errno.EMFILE


async def test_serve_listeners_connection_nursery(autojump_clock: MockClock) -> None:
    listener = MemoryListener()

    async def handler(stream: Stream) -> None:
        await trio.sleep(1)

    class Done(Exception):
        pass

    async def connection_watcher(
        *,
        task_status: TaskStatus[Nursery] = trio.TASK_STATUS_IGNORED,
    ) -> NoReturn:
        async with trio.open_nursery() as nursery:
            task_status.started(nursery)
            await wait_all_tasks_blocked()
            assert len(nursery.child_tasks) == 10
            raise Done

    # the exception is wrapped twice because we open two nested nurseries
    with RaisesGroup(RaisesGroup(Done)):
        async with trio.open_nursery() as nursery:
            value = await nursery.start(connection_watcher)
            assert isinstance(value, trio.Nursery)
            handler_nursery: trio.Nursery = value
            await nursery.start(
                partial(
                    trio.serve_listeners,
                    handler,
                    [listener],
                    handler_nursery=handler_nursery,
                ),
            )
            for _ in range(10):
                nursery.start_soon(listener.connect)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\wsproto\connection.py ===
"""
wsproto/connection
~~~~~~~~~~~~~~~~~~

An implementation of a WebSocket connection.
"""

from collections import deque
from enum import Enum
from typing import Deque, Generator, List, Optional

from .events import (
    BytesMessage,
    CloseConnection,
    Event,
    Message,
    Ping,
    Pong,
    TextMessage,
)
from .extensions import Extension
from .frame_protocol import CloseReason, FrameProtocol, Opcode, ParseFailed
from .utilities import LocalProtocolError


class ConnectionState(Enum):
    """
    RFC 6455, Section 4 - Opening Handshake
    """

    #: The opening handshake is in progress.
    CONNECTING = 0
    #: The opening handshake is complete.
    OPEN = 1
    #: The remote WebSocket has initiated a connection close.
    REMOTE_CLOSING = 2
    #: The local WebSocket (i.e. this instance) has initiated a connection close.
    LOCAL_CLOSING = 3
    #: The closing handshake has completed.
    CLOSED = 4
    #: The connection was rejected during the opening handshake.
    REJECTING = 5


class ConnectionType(Enum):
    """An enumeration of connection types."""

    #: This connection will act as client and talk to a remote server
    CLIENT = 1

    #: This connection will as as server and waits for client connections
    SERVER = 2


CLIENT = ConnectionType.CLIENT
SERVER = ConnectionType.SERVER


class Connection:
    """
    A low-level WebSocket connection object.

    This wraps two other protocol objects, an HTTP/1.1 protocol object used
    to do the initial HTTP upgrade handshake and a WebSocket frame protocol
    object used to exchange messages and other control frames.

    :param conn_type: Whether this object is on the client- or server-side of
        a connection. To initialise as a client pass ``CLIENT`` otherwise
        pass ``SERVER``.
    :type conn_type: ``ConnectionType``
    """

    def __init__(
        self,
        connection_type: ConnectionType,
        extensions: Optional[List[Extension]] = None,
        trailing_data: bytes = b"",
    ) -> None:
        self.client = connection_type is ConnectionType.CLIENT
        self._events: Deque[Event] = deque()
        self._proto = FrameProtocol(self.client, extensions or [])
        self._state = ConnectionState.OPEN
        self.receive_data(trailing_data)

    @property
    def state(self) -> ConnectionState:
        return self._state

    def send(self, event: Event) -> bytes:
        data = b""
        if isinstance(event, Message) and self.state == ConnectionState.OPEN:
            data += self._proto.send_data(event.data, event.message_finished)
        elif isinstance(event, Ping) and self.state == ConnectionState.OPEN:
            data += self._proto.ping(event.payload)
        elif isinstance(event, Pong) and self.state == ConnectionState.OPEN:
            data += self._proto.pong(event.payload)
        elif isinstance(event, CloseConnection) and self.state in {
            ConnectionState.OPEN,
            ConnectionState.REMOTE_CLOSING,
        }:
            data += self._proto.close(event.code, event.reason)
            if self.state == ConnectionState.REMOTE_CLOSING:
                self._state = ConnectionState.CLOSED
            else:
                self._state = ConnectionState.LOCAL_CLOSING
        else:
            raise LocalProtocolError(
                f"Event {event} cannot be sent in state {self.state}."
            )
        return data

    def receive_data(self, data: Optional[bytes]) -> None:
        """
        Pass some received data to the connection for handling.

        A list of events that the remote peer triggered by sending this data can
        be retrieved with :meth:`~wsproto.connection.Connection.events`.

        :param data: The data received from the remote peer on the network.
        :type data: ``bytes``
        """

        if data is None:
            # "If _The WebSocket Connection is Closed_ and no Close control
            # frame was received by the endpoint (such as could occur if the
            # underlying transport connection is lost), _The WebSocket
            # Connection Close Code_ is considered to be 1006."
            self._events.append(CloseConnection(code=CloseReason.ABNORMAL_CLOSURE))
            self._state = ConnectionState.CLOSED
            return

        if self.state in (ConnectionState.OPEN, ConnectionState.LOCAL_CLOSING):
            self._proto.receive_bytes(data)
        elif self.state is ConnectionState.CLOSED:
            raise LocalProtocolError("Connection already closed.")
        else:
            pass  # pragma: no cover

    def events(self) -> Generator[Event, None, None]:
        """
        Return a generator that provides any events that have been generated
        by protocol activity.

        :returns: generator of :class:`Event <wsproto.events.Event>` subclasses
        """
        while self._events:
            yield self._events.popleft()

        try:
            for frame in self._proto.received_frames():
                if frame.opcode is Opcode.PING:
                    assert frame.frame_finished and frame.message_finished
                    assert isinstance(frame.payload, (bytes, bytearray))
                    yield Ping(payload=frame.payload)

                elif frame.opcode is Opcode.PONG:
                    assert frame.frame_finished and frame.message_finished
                    assert isinstance(frame.payload, (bytes, bytearray))
                    yield Pong(payload=frame.payload)

                elif frame.opcode is Opcode.CLOSE:
                    assert isinstance(frame.payload, tuple)
                    code, reason = frame.payload
                    if self.state is ConnectionState.LOCAL_CLOSING:
                        self._state = ConnectionState.CLOSED
                    else:
                        self._state = ConnectionState.REMOTE_CLOSING
                    yield CloseConnection(code=code, reason=reason)

                elif frame.opcode is Opcode.TEXT:
                    assert isinstance(frame.payload, str)
                    yield TextMessage(
                        data=frame.payload,
                        frame_finished=frame.frame_finished,
                        message_finished=frame.message_finished,
                    )

                elif frame.opcode is Opcode.BINARY:
                    assert isinstance(frame.payload, (bytes, bytearray))
                    yield BytesMessage(
                        data=frame.payload,
                        frame_finished=frame.frame_finished,
                        message_finished=frame.message_finished,
                    )

                else:
                    pass  # pragma: no cover
        except ParseFailed as exc:
            yield CloseConnection(code=exc.code, reason=str(exc))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v135\log.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: Log
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing
from . import network
from . import runtime


@dataclass
class LogEntry:
    '''
    Log entry.
    '''
    #: Log entry source.
    source: str

    #: Log entry severity.
    level: str

    #: Logged text.
    text: str

    #: Timestamp when this entry was added.
    timestamp: runtime.Timestamp

    category: typing.Optional[str] = None

    #: URL of the resource if known.
    url: typing.Optional[str] = None

    #: Line number in the resource.
    line_number: typing.Optional[int] = None

    #: JavaScript stack trace.
    stack_trace: typing.Optional[runtime.StackTrace] = None

    #: Identifier of the network request associated with this entry.
    network_request_id: typing.Optional[network.RequestId] = None

    #: Identifier of the worker associated with this entry.
    worker_id: typing.Optional[str] = None

    #: Call arguments.
    args: typing.Optional[typing.List[runtime.RemoteObject]] = None

    def to_json(self):
        json = dict()
        json['source'] = self.source
        json['level'] = self.level
        json['text'] = self.text
        json['timestamp'] = self.timestamp.to_json()
        if self.category is not None:
            json['category'] = self.category
        if self.url is not None:
            json['url'] = self.url
        if self.line_number is not None:
            json['lineNumber'] = self.line_number
        if self.stack_trace is not None:
            json['stackTrace'] = self.stack_trace.to_json()
        if self.network_request_id is not None:
            json['networkRequestId'] = self.network_request_id.to_json()
        if self.worker_id is not None:
            json['workerId'] = self.worker_id
        if self.args is not None:
            json['args'] = [i.to_json() for i in self.args]
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            source=str(json['source']),
            level=str(json['level']),
            text=str(json['text']),
            timestamp=runtime.Timestamp.from_json(json['timestamp']),
            category=str(json['category']) if 'category' in json else None,
            url=str(json['url']) if 'url' in json else None,
            line_number=int(json['lineNumber']) if 'lineNumber' in json else None,
            stack_trace=runtime.StackTrace.from_json(json['stackTrace']) if 'stackTrace' in json else None,
            network_request_id=network.RequestId.from_json(json['networkRequestId']) if 'networkRequestId' in json else None,
            worker_id=str(json['workerId']) if 'workerId' in json else None,
            args=[runtime.RemoteObject.from_json(i) for i in json['args']] if 'args' in json else None,
        )


@dataclass
class ViolationSetting:
    '''
    Violation configuration setting.
    '''
    #: Violation type.
    name: str

    #: Time threshold to trigger upon.
    threshold: float

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['threshold'] = self.threshold
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            threshold=float(json['threshold']),
        )


def clear() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Clears the log.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.clear',
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disables log domain, prevents further log entries from being reported to the client.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.disable',
    }
    json = yield cmd_dict


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enables log domain, sends the entries collected so far to the client by means of the
    ``entryAdded`` notification.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.enable',
    }
    json = yield cmd_dict


def start_violations_report(
        config: typing.List[ViolationSetting]
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    start violation reporting.

    :param config: Configuration for violations.
    '''
    params: T_JSON_DICT = dict()
    params['config'] = [i.to_json() for i in config]
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.startViolationsReport',
        'params': params,
    }
    json = yield cmd_dict


def stop_violations_report() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stop violation reporting.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.stopViolationsReport',
    }
    json = yield cmd_dict


@event_class('Log.entryAdded')
@dataclass
class EntryAdded:
    '''
    Issued when new message was logged.
    '''
    #: The entry.
    entry: LogEntry

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> EntryAdded:
        return cls(
            entry=LogEntry.from_json(json['entry'])
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v136\log.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: Log
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing
from . import network
from . import runtime


@dataclass
class LogEntry:
    '''
    Log entry.
    '''
    #: Log entry source.
    source: str

    #: Log entry severity.
    level: str

    #: Logged text.
    text: str

    #: Timestamp when this entry was added.
    timestamp: runtime.Timestamp

    category: typing.Optional[str] = None

    #: URL of the resource if known.
    url: typing.Optional[str] = None

    #: Line number in the resource.
    line_number: typing.Optional[int] = None

    #: JavaScript stack trace.
    stack_trace: typing.Optional[runtime.StackTrace] = None

    #: Identifier of the network request associated with this entry.
    network_request_id: typing.Optional[network.RequestId] = None

    #: Identifier of the worker associated with this entry.
    worker_id: typing.Optional[str] = None

    #: Call arguments.
    args: typing.Optional[typing.List[runtime.RemoteObject]] = None

    def to_json(self):
        json = dict()
        json['source'] = self.source
        json['level'] = self.level
        json['text'] = self.text
        json['timestamp'] = self.timestamp.to_json()
        if self.category is not None:
            json['category'] = self.category
        if self.url is not None:
            json['url'] = self.url
        if self.line_number is not None:
            json['lineNumber'] = self.line_number
        if self.stack_trace is not None:
            json['stackTrace'] = self.stack_trace.to_json()
        if self.network_request_id is not None:
            json['networkRequestId'] = self.network_request_id.to_json()
        if self.worker_id is not None:
            json['workerId'] = self.worker_id
        if self.args is not None:
            json['args'] = [i.to_json() for i in self.args]
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            source=str(json['source']),
            level=str(json['level']),
            text=str(json['text']),
            timestamp=runtime.Timestamp.from_json(json['timestamp']),
            category=str(json['category']) if 'category' in json else None,
            url=str(json['url']) if 'url' in json else None,
            line_number=int(json['lineNumber']) if 'lineNumber' in json else None,
            stack_trace=runtime.StackTrace.from_json(json['stackTrace']) if 'stackTrace' in json else None,
            network_request_id=network.RequestId.from_json(json['networkRequestId']) if 'networkRequestId' in json else None,
            worker_id=str(json['workerId']) if 'workerId' in json else None,
            args=[runtime.RemoteObject.from_json(i) for i in json['args']] if 'args' in json else None,
        )


@dataclass
class ViolationSetting:
    '''
    Violation configuration setting.
    '''
    #: Violation type.
    name: str

    #: Time threshold to trigger upon.
    threshold: float

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['threshold'] = self.threshold
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            threshold=float(json['threshold']),
        )


def clear() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Clears the log.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.clear',
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disables log domain, prevents further log entries from being reported to the client.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.disable',
    }
    json = yield cmd_dict


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enables log domain, sends the entries collected so far to the client by means of the
    ``entryAdded`` notification.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.enable',
    }
    json = yield cmd_dict


def start_violations_report(
        config: typing.List[ViolationSetting]
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    start violation reporting.

    :param config: Configuration for violations.
    '''
    params: T_JSON_DICT = dict()
    params['config'] = [i.to_json() for i in config]
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.startViolationsReport',
        'params': params,
    }
    json = yield cmd_dict


def stop_violations_report() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stop violation reporting.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.stopViolationsReport',
    }
    json = yield cmd_dict


@event_class('Log.entryAdded')
@dataclass
class EntryAdded:
    '''
    Issued when new message was logged.
    '''
    #: The entry.
    entry: LogEntry

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> EntryAdded:
        return cls(
            entry=LogEntry.from_json(json['entry'])
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v137\log.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: Log
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing
from . import network
from . import runtime


@dataclass
class LogEntry:
    '''
    Log entry.
    '''
    #: Log entry source.
    source: str

    #: Log entry severity.
    level: str

    #: Logged text.
    text: str

    #: Timestamp when this entry was added.
    timestamp: runtime.Timestamp

    category: typing.Optional[str] = None

    #: URL of the resource if known.
    url: typing.Optional[str] = None

    #: Line number in the resource.
    line_number: typing.Optional[int] = None

    #: JavaScript stack trace.
    stack_trace: typing.Optional[runtime.StackTrace] = None

    #: Identifier of the network request associated with this entry.
    network_request_id: typing.Optional[network.RequestId] = None

    #: Identifier of the worker associated with this entry.
    worker_id: typing.Optional[str] = None

    #: Call arguments.
    args: typing.Optional[typing.List[runtime.RemoteObject]] = None

    def to_json(self):
        json = dict()
        json['source'] = self.source
        json['level'] = self.level
        json['text'] = self.text
        json['timestamp'] = self.timestamp.to_json()
        if self.category is not None:
            json['category'] = self.category
        if self.url is not None:
            json['url'] = self.url
        if self.line_number is not None:
            json['lineNumber'] = self.line_number
        if self.stack_trace is not None:
            json['stackTrace'] = self.stack_trace.to_json()
        if self.network_request_id is not None:
            json['networkRequestId'] = self.network_request_id.to_json()
        if self.worker_id is not None:
            json['workerId'] = self.worker_id
        if self.args is not None:
            json['args'] = [i.to_json() for i in self.args]
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            source=str(json['source']),
            level=str(json['level']),
            text=str(json['text']),
            timestamp=runtime.Timestamp.from_json(json['timestamp']),
            category=str(json['category']) if 'category' in json else None,
            url=str(json['url']) if 'url' in json else None,
            line_number=int(json['lineNumber']) if 'lineNumber' in json else None,
            stack_trace=runtime.StackTrace.from_json(json['stackTrace']) if 'stackTrace' in json else None,
            network_request_id=network.RequestId.from_json(json['networkRequestId']) if 'networkRequestId' in json else None,
            worker_id=str(json['workerId']) if 'workerId' in json else None,
            args=[runtime.RemoteObject.from_json(i) for i in json['args']] if 'args' in json else None,
        )


@dataclass
class ViolationSetting:
    '''
    Violation configuration setting.
    '''
    #: Violation type.
    name: str

    #: Time threshold to trigger upon.
    threshold: float

    def to_json(self):
        json = dict()
        json['name'] = self.name
        json['threshold'] = self.threshold
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            name=str(json['name']),
            threshold=float(json['threshold']),
        )


def clear() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Clears the log.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.clear',
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disables log domain, prevents further log entries from being reported to the client.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.disable',
    }
    json = yield cmd_dict


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enables log domain, sends the entries collected so far to the client by means of the
    ``entryAdded`` notification.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.enable',
    }
    json = yield cmd_dict


def start_violations_report(
        config: typing.List[ViolationSetting]
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    start violation reporting.

    :param config: Configuration for violations.
    '''
    params: T_JSON_DICT = dict()
    params['config'] = [i.to_json() for i in config]
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.startViolationsReport',
        'params': params,
    }
    json = yield cmd_dict


def stop_violations_report() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Stop violation reporting.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'Log.stopViolationsReport',
    }
    json = yield cmd_dict


@event_class('Log.entryAdded')
@dataclass
class EntryAdded:
    '''
    Issued when new message was logged.
    '''
    #: The entry.
    entry: LogEntry

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> EntryAdded:
        return cls(
            entry=LogEntry.from_json(json['entry'])
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\ntheory\multinomial.py ===
from sympy.utilities.misc import as_int


def binomial_coefficients(n):
    """Return a dictionary containing pairs :math:`{(k1,k2) : C_kn}` where
    :math:`C_kn` are binomial coefficients and :math:`n=k1+k2`.

    Examples
    ========

    >>> from sympy.ntheory import binomial_coefficients
    >>> binomial_coefficients(9)
    {(0, 9): 1, (1, 8): 9, (2, 7): 36, (3, 6): 84,
     (4, 5): 126, (5, 4): 126, (6, 3): 84, (7, 2): 36, (8, 1): 9, (9, 0): 1}

    See Also
    ========

    binomial_coefficients_list, multinomial_coefficients
    """
    n = as_int(n)
    d = {(0, n): 1, (n, 0): 1}
    a = 1
    for k in range(1, n//2 + 1):
        a = (a * (n - k + 1))//k
        d[k, n - k] = d[n - k, k] = a
    return d


def binomial_coefficients_list(n):
    """ Return a list of binomial coefficients as rows of the Pascal's
    triangle.

    Examples
    ========

    >>> from sympy.ntheory import binomial_coefficients_list
    >>> binomial_coefficients_list(9)
    [1, 9, 36, 84, 126, 126, 84, 36, 9, 1]

    See Also
    ========

    binomial_coefficients, multinomial_coefficients
    """
    n = as_int(n)
    d = [1] * (n + 1)
    a = 1
    for k in range(1, n//2 + 1):
        a = (a * (n - k + 1))//k
        d[k] = d[n - k] = a
    return d


def multinomial_coefficients(m, n):
    r"""Return a dictionary containing pairs ``{(k1,k2,..,km) : C_kn}``
    where ``C_kn`` are multinomial coefficients such that
    ``n=k1+k2+..+km``.

    Examples
    ========

    >>> from sympy.ntheory import multinomial_coefficients
    >>> multinomial_coefficients(2, 5) # indirect doctest
    {(0, 5): 1, (1, 4): 5, (2, 3): 10, (3, 2): 10, (4, 1): 5, (5, 0): 1}

    Notes
    =====

    The algorithm is based on the following result:

    .. math::
        \binom{n}{k_1, \ldots, k_m} =
        \frac{k_1 + 1}{n - k_1} \sum_{i=2}^m \binom{n}{k_1 + 1, \ldots, k_i - 1, \ldots}

    Code contributed to Sage by Yann Laigle-Chapuy, copied with permission
    of the author.

    See Also
    ========

    binomial_coefficients_list, binomial_coefficients
    """
    m = as_int(m)
    n = as_int(n)
    if not m:
        if n:
            return {}
        return {(): 1}
    if m == 2:
        return binomial_coefficients(n)
    if m >= 2*n and n > 1:
        return dict(multinomial_coefficients_iterator(m, n))
    t = [n] + [0] * (m - 1)
    r = {tuple(t): 1}
    if n:
        j = 0  # j will be the leftmost nonzero position
    else:
        j = m
    # enumerate tuples in co-lex order
    while j < m - 1:
        # compute next tuple
        tj = t[j]
        if j:
            t[j] = 0
            t[0] = tj
        if tj > 1:
            t[j + 1] += 1
            j = 0
            start = 1
            v = 0
        else:
            j += 1
            start = j + 1
            v = r[tuple(t)]
            t[j] += 1
        # compute the value
        # NB: the initialization of v was done above
        for k in range(start, m):
            if t[k]:
                t[k] -= 1
                v += r[tuple(t)]
                t[k] += 1
        t[0] -= 1
        r[tuple(t)] = (v * tj) // (n - t[0])
    return r


def multinomial_coefficients_iterator(m, n, _tuple=tuple):
    """multinomial coefficient iterator

    This routine has been optimized for `m` large with respect to `n` by taking
    advantage of the fact that when the monomial tuples `t` are stripped of
    zeros, their coefficient is the same as that of the monomial tuples from
    ``multinomial_coefficients(n, n)``. Therefore, the latter coefficients are
    precomputed to save memory and time.

    >>> from sympy.ntheory.multinomial import multinomial_coefficients
    >>> m53, m33 = multinomial_coefficients(5,3), multinomial_coefficients(3,3)
    >>> m53[(0,0,0,1,2)] == m53[(0,0,1,0,2)] == m53[(1,0,2,0,0)] == m33[(0,1,2)]
    True

    Examples
    ========

    >>> from sympy.ntheory.multinomial import multinomial_coefficients_iterator
    >>> it = multinomial_coefficients_iterator(20,3)
    >>> next(it)
    ((3, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0), 1)
    """
    m = as_int(m)
    n = as_int(n)
    if m < 2*n or n == 1:
        mc = multinomial_coefficients(m, n)
        yield from mc.items()
    else:
        mc = multinomial_coefficients(n, n)
        mc1 = {}
        for k, v in mc.items():
            mc1[_tuple(filter(None, k))] = v
        mc = mc1

        t = [n] + [0] * (m - 1)
        t1 = _tuple(t)
        b = _tuple(filter(None, t1))
        yield (t1, mc[b])
        if n:
            j = 0  # j will be the leftmost nonzero position
        else:
            j = m
        # enumerate tuples in co-lex order
        while j < m - 1:
            # compute next tuple
            tj = t[j]
            if j:
                t[j] = 0
                t[0] = tj
            if tj > 1:
                t[j + 1] += 1
                j = 0
            else:
                j += 1
                t[j] += 1

            t[0] -= 1
            t1 = _tuple(t)
            b = _tuple(filter(None, t1))
            yield (t1, mc[b])

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\plotting\plotgrid.py ===

from sympy.external import import_module
import sympy.plotting.backends.base_backend as base_backend


# N.B.
# When changing the minimum module version for matplotlib, please change
# the same in the `SymPyDocTestFinder`` in `sympy/testing/runtests.py`


__doctest_requires__ = {
    ("PlotGrid",): ["matplotlib"],
}


class PlotGrid:
    """This class helps to plot subplots from already created SymPy plots
    in a single figure.

    Examples
    ========

    .. plot::
       :context: close-figs
       :format: doctest
       :include-source: True

        >>> from sympy import symbols
        >>> from sympy.plotting import plot, plot3d, PlotGrid
        >>> x, y = symbols('x, y')
        >>> p1 = plot(x, x**2, x**3, (x, -5, 5))
        >>> p2 = plot((x**2, (x, -6, 6)), (x, (x, -5, 5)))
        >>> p3 = plot(x**3, (x, -5, 5))
        >>> p4 = plot3d(x*y, (x, -5, 5), (y, -5, 5))

    Plotting vertically in a single line:

    .. plot::
       :context: close-figs
       :format: doctest
       :include-source: True

        >>> PlotGrid(2, 1, p1, p2)
        PlotGrid object containing:
        Plot[0]:Plot object containing:
        [0]: cartesian line: x for x over (-5.0, 5.0)
        [1]: cartesian line: x**2 for x over (-5.0, 5.0)
        [2]: cartesian line: x**3 for x over (-5.0, 5.0)
        Plot[1]:Plot object containing:
        [0]: cartesian line: x**2 for x over (-6.0, 6.0)
        [1]: cartesian line: x for x over (-5.0, 5.0)

    Plotting horizontally in a single line:

    .. plot::
       :context: close-figs
       :format: doctest
       :include-source: True

        >>> PlotGrid(1, 3, p2, p3, p4)
        PlotGrid object containing:
        Plot[0]:Plot object containing:
        [0]: cartesian line: x**2 for x over (-6.0, 6.0)
        [1]: cartesian line: x for x over (-5.0, 5.0)
        Plot[1]:Plot object containing:
        [0]: cartesian line: x**3 for x over (-5.0, 5.0)
        Plot[2]:Plot object containing:
        [0]: cartesian surface: x*y for x over (-5.0, 5.0) and y over (-5.0, 5.0)

    Plotting in a grid form:

    .. plot::
       :context: close-figs
       :format: doctest
       :include-source: True

        >>> PlotGrid(2, 2, p1, p2, p3, p4)
        PlotGrid object containing:
        Plot[0]:Plot object containing:
        [0]: cartesian line: x for x over (-5.0, 5.0)
        [1]: cartesian line: x**2 for x over (-5.0, 5.0)
        [2]: cartesian line: x**3 for x over (-5.0, 5.0)
        Plot[1]:Plot object containing:
        [0]: cartesian line: x**2 for x over (-6.0, 6.0)
        [1]: cartesian line: x for x over (-5.0, 5.0)
        Plot[2]:Plot object containing:
        [0]: cartesian line: x**3 for x over (-5.0, 5.0)
        Plot[3]:Plot object containing:
        [0]: cartesian surface: x*y for x over (-5.0, 5.0) and y over (-5.0, 5.0)

    """
    def __init__(self, nrows, ncolumns, *args, show=True, size=None, **kwargs):
        """
        Parameters
        ==========

        nrows :
            The number of rows that should be in the grid of the
            required subplot.
        ncolumns :
            The number of columns that should be in the grid
            of the required subplot.

        nrows and ncolumns together define the required grid.

        Arguments
        =========

        A list of predefined plot objects entered in a row-wise sequence
        i.e. plot objects which are to be in the top row of the required
        grid are written first, then the second row objects and so on

        Keyword arguments
        =================

        show : Boolean
            The default value is set to ``True``. Set show to ``False`` and
            the function will not display the subplot. The returned instance
            of the ``PlotGrid`` class can then be used to save or display the
            plot by calling the ``save()`` and ``show()`` methods
            respectively.
        size : (float, float), optional
            A tuple in the form (width, height) in inches to specify the size of
            the overall figure. The default value is set to ``None``, meaning
            the size will be set by the default backend.
        """
        self.matplotlib = import_module('matplotlib',
            import_kwargs={'fromlist': ['pyplot', 'cm', 'collections']},
            min_module_version='1.1.0', catch=(RuntimeError,))
        self.nrows = nrows
        self.ncolumns = ncolumns
        self._series = []
        self._fig = None
        self.args = args
        for arg in args:
            self._series.append(arg._series)
        self.size = size
        if show and self.matplotlib:
            self.show()

    def _create_figure(self):
        gs = self.matplotlib.gridspec.GridSpec(self.nrows, self.ncolumns)
        mapping = {}
        c = 0
        for i in range(self.nrows):
            for j in range(self.ncolumns):
                if c < len(self.args):
                    mapping[gs[i, j]] = self.args[c]
                c += 1

        kw = {} if not self.size else {"figsize": self.size}
        self._fig = self.matplotlib.pyplot.figure(**kw)
        for spec, p in mapping.items():
            kw = ({"projection": "3d"} if (len(p._series) > 0 and
                p._series[0].is_3D) else {})
            cur_ax = self._fig.add_subplot(spec, **kw)
            p._plotgrid_fig = self._fig
            p._plotgrid_ax = cur_ax
            p.process_series()

    @property
    def fig(self):
        if not self._fig:
            self._create_figure()
        return self._fig

    @property
    def _backend(self):
        return self

    def close(self):
        self.matplotlib.pyplot.close(self.fig)

    def show(self):
        if base_backend._show:
            self.fig.tight_layout()
            self.matplotlib.pyplot.show()
        else:
            self.close()

    def save(self, path):
        self.fig.savefig(path)

    def __str__(self):
        plot_strs = [('Plot[%d]:' % i) + str(plot)
                      for i, plot in enumerate(self.args)]

        return 'PlotGrid object containing:\n' + '\n'.join(plot_strs)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\plotting\pygletplot\util.py ===
try:
    from ctypes import c_float, c_int, c_double
except ImportError:
    pass

import pyglet.gl as pgl
from sympy.core import S


def get_model_matrix(array_type=c_float, glGetMethod=pgl.glGetFloatv):
    """
    Returns the current modelview matrix.
    """
    m = (array_type*16)()
    glGetMethod(pgl.GL_MODELVIEW_MATRIX, m)
    return m


def get_projection_matrix(array_type=c_float, glGetMethod=pgl.glGetFloatv):
    """
    Returns the current modelview matrix.
    """
    m = (array_type*16)()
    glGetMethod(pgl.GL_PROJECTION_MATRIX, m)
    return m


def get_viewport():
    """
    Returns the current viewport.
    """
    m = (c_int*4)()
    pgl.glGetIntegerv(pgl.GL_VIEWPORT, m)
    return m


def get_direction_vectors():
    m = get_model_matrix()
    return ((m[0], m[4], m[8]),
            (m[1], m[5], m[9]),
            (m[2], m[6], m[10]))


def get_view_direction_vectors():
    m = get_model_matrix()
    return ((m[0], m[1], m[2]),
            (m[4], m[5], m[6]),
            (m[8], m[9], m[10]))


def get_basis_vectors():
    return ((1, 0, 0), (0, 1, 0), (0, 0, 1))


def screen_to_model(x, y, z):
    m = get_model_matrix(c_double, pgl.glGetDoublev)
    p = get_projection_matrix(c_double, pgl.glGetDoublev)
    w = get_viewport()
    mx, my, mz = c_double(), c_double(), c_double()
    pgl.gluUnProject(x, y, z, m, p, w, mx, my, mz)
    return float(mx.value), float(my.value), float(mz.value)


def model_to_screen(x, y, z):
    m = get_model_matrix(c_double, pgl.glGetDoublev)
    p = get_projection_matrix(c_double, pgl.glGetDoublev)
    w = get_viewport()
    mx, my, mz = c_double(), c_double(), c_double()
    pgl.gluProject(x, y, z, m, p, w, mx, my, mz)
    return float(mx.value), float(my.value), float(mz.value)


def vec_subs(a, b):
    return tuple(a[i] - b[i] for i in range(len(a)))


def billboard_matrix():
    """
    Removes rotational components of
    current matrix so that primitives
    are always drawn facing the viewer.

    |1|0|0|x|
    |0|1|0|x|
    |0|0|1|x| (x means left unchanged)
    |x|x|x|x|
    """
    m = get_model_matrix()
    # XXX: for i in range(11): m[i] = i ?
    m[0] = 1
    m[1] = 0
    m[2] = 0
    m[4] = 0
    m[5] = 1
    m[6] = 0
    m[8] = 0
    m[9] = 0
    m[10] = 1
    pgl.glLoadMatrixf(m)


def create_bounds():
    return [[S.Infinity, S.NegativeInfinity, 0],
            [S.Infinity, S.NegativeInfinity, 0],
            [S.Infinity, S.NegativeInfinity, 0]]


def update_bounds(b, v):
    if v is None:
        return
    for axis in range(3):
        b[axis][0] = min([b[axis][0], v[axis]])
        b[axis][1] = max([b[axis][1], v[axis]])


def interpolate(a_min, a_max, a_ratio):
    return a_min + a_ratio * (a_max - a_min)


def rinterpolate(a_min, a_max, a_value):
    a_range = a_max - a_min
    if a_max == a_min:
        a_range = 1.0
    return (a_value - a_min) / float(a_range)


def interpolate_color(color1, color2, ratio):
    return tuple(interpolate(color1[i], color2[i], ratio) for i in range(3))


def scale_value(v, v_min, v_len):
    return (v - v_min) / v_len


def scale_value_list(flist):
    v_min, v_max = min(flist), max(flist)
    v_len = v_max - v_min
    return [scale_value(f, v_min, v_len) for f in flist]


def strided_range(r_min, r_max, stride, max_steps=50):
    o_min, o_max = r_min, r_max
    if abs(r_min - r_max) < 0.001:
        return []
    try:
        range(int(r_min - r_max))
    except (TypeError, OverflowError):
        return []
    if r_min > r_max:
        raise ValueError("r_min cannot be greater than r_max")
    r_min_s = (r_min % stride)
    r_max_s = stride - (r_max % stride)
    if abs(r_max_s - stride) < 0.001:
        r_max_s = 0.0
    r_min -= r_min_s
    r_max += r_max_s
    r_steps = int((r_max - r_min)/stride)
    if max_steps and r_steps > max_steps:
        return strided_range(o_min, o_max, stride*2)
    return [r_min] + [r_min + e*stride for e in range(1, r_steps + 1)] + [r_max]


def parse_option_string(s):
    if not isinstance(s, str):
        return None
    options = {}
    for token in s.split(';'):
        pieces = token.split('=')
        if len(pieces) == 1:
            option, value = pieces[0], ""
        elif len(pieces) == 2:
            option, value = pieces
        else:
            raise ValueError("Plot option string '%s' is malformed." % (s))
        options[option.strip()] = value.strip()
    return options


def dot_product(v1, v2):
    return sum(v1[i]*v2[i] for i in range(3))


def vec_sub(v1, v2):
    return tuple(v1[i] - v2[i] for i in range(3))


def vec_mag(v):
    return sum(v[i]**2 for i in range(3))**(0.5)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\domains\old_fractionfield.py ===
"""Implementation of :class:`FractionField` class. """


from sympy.polys.domains.field import Field
from sympy.polys.domains.compositedomain import CompositeDomain
from sympy.polys.polyclasses import DMF
from sympy.polys.polyerrors import GeneratorsNeeded
from sympy.polys.polyutils import dict_from_basic, basic_from_dict, _dict_reorder
from sympy.utilities import public

@public
class FractionField(Field, CompositeDomain):
    """A class for representing rational function fields. """

    dtype = DMF
    is_FractionField = is_Frac = True

    has_assoc_Ring = True
    has_assoc_Field = True

    def __init__(self, dom, *gens):
        if not gens:
            raise GeneratorsNeeded("generators not specified")

        lev = len(gens) - 1
        self.ngens = len(gens)

        self.zero = self.dtype.zero(lev, dom)
        self.one = self.dtype.one(lev, dom)

        self.domain = self.dom = dom
        self.symbols = self.gens = gens

    def set_domain(self, dom):
        """Make a new fraction field with given domain. """
        return self.__class__(dom, *self.gens)

    def new(self, element):
        return self.dtype(element, self.dom, len(self.gens) - 1)

    def __str__(self):
        return str(self.dom) + '(' + ','.join(map(str, self.gens)) + ')'

    def __hash__(self):
        return hash((self.__class__.__name__, self.dtype, self.dom, self.gens))

    def __eq__(self, other):
        """Returns ``True`` if two domains are equivalent. """
        return isinstance(other, FractionField) and \
            self.dtype == other.dtype and self.dom == other.dom and self.gens == other.gens

    def to_sympy(self, a):
        """Convert ``a`` to a SymPy object. """
        return (basic_from_dict(a.numer().to_sympy_dict(), *self.gens) /
                basic_from_dict(a.denom().to_sympy_dict(), *self.gens))

    def from_sympy(self, a):
        """Convert SymPy's expression to ``dtype``. """
        p, q = a.as_numer_denom()

        num, _ = dict_from_basic(p, gens=self.gens)
        den, _ = dict_from_basic(q, gens=self.gens)

        for k, v in num.items():
            num[k] = self.dom.from_sympy(v)

        for k, v in den.items():
            den[k] = self.dom.from_sympy(v)

        return self((num, den)).cancel()

    def from_ZZ(K1, a, K0):
        """Convert a Python ``int`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_ZZ_python(K1, a, K0):
        """Convert a Python ``int`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_QQ_python(K1, a, K0):
        """Convert a Python ``Fraction`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_ZZ_gmpy(K1, a, K0):
        """Convert a GMPY ``mpz`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_QQ_gmpy(K1, a, K0):
        """Convert a GMPY ``mpq`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_RealField(K1, a, K0):
        """Convert a mpmath ``mpf`` object to ``dtype``. """
        return K1(K1.dom.convert(a, K0))

    def from_GlobalPolynomialRing(K1, a, K0):
        """Convert a ``DMF`` object to ``dtype``. """
        if K1.gens == K0.gens:
            if K1.dom == K0.dom:
                return K1(a.to_list())
            else:
                return K1(a.convert(K1.dom).to_list())
        else:
            monoms, coeffs = _dict_reorder(a.to_dict(), K0.gens, K1.gens)

            if K1.dom != K0.dom:
                coeffs = [ K1.dom.convert(c, K0.dom) for c in coeffs ]

            return K1(dict(zip(monoms, coeffs)))

    def from_FractionField(K1, a, K0):
        """
        Convert a fraction field element to another fraction field.

        Examples
        ========

        >>> from sympy.polys.polyclasses import DMF
        >>> from sympy.polys.domains import ZZ, QQ
        >>> from sympy.abc import x

        >>> f = DMF(([ZZ(1), ZZ(2)], [ZZ(1), ZZ(1)]), ZZ)

        >>> QQx = QQ.old_frac_field(x)
        >>> ZZx = ZZ.old_frac_field(x)

        >>> QQx.from_FractionField(f, ZZx)
        DMF([1, 2], [1, 1], QQ)

        """
        if K1.gens == K0.gens:
            if K1.dom == K0.dom:
                return a
            else:
                return K1((a.numer().convert(K1.dom).to_list(),
                           a.denom().convert(K1.dom).to_list()))
        elif set(K0.gens).issubset(K1.gens):
            nmonoms, ncoeffs = _dict_reorder(
                a.numer().to_dict(), K0.gens, K1.gens)
            dmonoms, dcoeffs = _dict_reorder(
                a.denom().to_dict(), K0.gens, K1.gens)

            if K1.dom != K0.dom:
                ncoeffs = [ K1.dom.convert(c, K0.dom) for c in ncoeffs ]
                dcoeffs = [ K1.dom.convert(c, K0.dom) for c in dcoeffs ]

            return K1((dict(zip(nmonoms, ncoeffs)), dict(zip(dmonoms, dcoeffs))))

    def get_ring(self):
        """Returns a ring associated with ``self``. """
        from sympy.polys.domains import PolynomialRing
        return PolynomialRing(self.dom, *self.gens)

    def poly_ring(self, *gens):
        """Returns a polynomial ring, i.e. `K[X]`. """
        raise NotImplementedError('nested domains not allowed')

    def frac_field(self, *gens):
        """Returns a fraction field, i.e. `K(X)`. """
        raise NotImplementedError('nested domains not allowed')

    def is_positive(self, a):
        """Returns True if ``a`` is positive. """
        return self.dom.is_positive(a.numer().LC())

    def is_negative(self, a):
        """Returns True if ``a`` is negative. """
        return self.dom.is_negative(a.numer().LC())

    def is_nonpositive(self, a):
        """Returns True if ``a`` is non-positive. """
        return self.dom.is_nonpositive(a.numer().LC())

    def is_nonnegative(self, a):
        """Returns True if ``a`` is non-negative. """
        return self.dom.is_nonnegative(a.numer().LC())

    def numer(self, a):
        """Returns numerator of ``a``. """
        return a.numer()

    def denom(self, a):
        """Returns denominator of ``a``. """
        return a.denom()

    def factorial(self, a):
        """Returns factorial of ``a``. """
        return self.dtype(self.dom.factorial(a))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\websocket\_socket.py ===
import errno
import selectors
import socket
from typing import Union

from ._exceptions import (
    WebSocketConnectionClosedException,
    WebSocketTimeoutException,
)
from ._ssl_compat import SSLError, SSLWantReadError, SSLWantWriteError
from ._utils import extract_error_code, extract_err_message

"""
_socket.py
websocket - WebSocket client library for Python

Copyright 2024 engn33r

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

DEFAULT_SOCKET_OPTION = [(socket.SOL_TCP, socket.TCP_NODELAY, 1)]
if hasattr(socket, "SO_KEEPALIVE"):
    DEFAULT_SOCKET_OPTION.append((socket.SOL_SOCKET, socket.SO_KEEPALIVE, 1))
if hasattr(socket, "TCP_KEEPIDLE"):
    DEFAULT_SOCKET_OPTION.append((socket.SOL_TCP, socket.TCP_KEEPIDLE, 30))
if hasattr(socket, "TCP_KEEPINTVL"):
    DEFAULT_SOCKET_OPTION.append((socket.SOL_TCP, socket.TCP_KEEPINTVL, 10))
if hasattr(socket, "TCP_KEEPCNT"):
    DEFAULT_SOCKET_OPTION.append((socket.SOL_TCP, socket.TCP_KEEPCNT, 3))

_default_timeout = None

__all__ = [
    "DEFAULT_SOCKET_OPTION",
    "sock_opt",
    "setdefaulttimeout",
    "getdefaulttimeout",
    "recv",
    "recv_line",
    "send",
]


class sock_opt:
    def __init__(self, sockopt: list, sslopt: dict) -> None:
        if sockopt is None:
            sockopt = []
        if sslopt is None:
            sslopt = {}
        self.sockopt = sockopt
        self.sslopt = sslopt
        self.timeout = None


def setdefaulttimeout(timeout: Union[int, float, None]) -> None:
    """
    Set the global timeout setting to connect.

    Parameters
    ----------
    timeout: int or float
        default socket timeout time (in seconds)
    """
    global _default_timeout
    _default_timeout = timeout


def getdefaulttimeout() -> Union[int, float, None]:
    """
    Get default timeout

    Returns
    ----------
    _default_timeout: int or float
        Return the global timeout setting (in seconds) to connect.
    """
    return _default_timeout


def recv(sock: socket.socket, bufsize: int) -> bytes:
    if not sock:
        raise WebSocketConnectionClosedException("socket is already closed.")

    def _recv():
        try:
            return sock.recv(bufsize)
        except SSLWantReadError:
            pass
        except socket.error as exc:
            error_code = extract_error_code(exc)
            if error_code not in [errno.EAGAIN, errno.EWOULDBLOCK]:
                raise

        sel = selectors.DefaultSelector()
        sel.register(sock, selectors.EVENT_READ)

        r = sel.select(sock.gettimeout())
        sel.close()

        if r:
            return sock.recv(bufsize)

    try:
        if sock.gettimeout() == 0:
            bytes_ = sock.recv(bufsize)
        else:
            bytes_ = _recv()
    except TimeoutError:
        raise WebSocketTimeoutException("Connection timed out")
    except socket.timeout as e:
        message = extract_err_message(e)
        raise WebSocketTimeoutException(message)
    except SSLError as e:
        message = extract_err_message(e)
        if isinstance(message, str) and "timed out" in message:
            raise WebSocketTimeoutException(message)
        else:
            raise

    if not bytes_:
        raise WebSocketConnectionClosedException("Connection to remote host was lost.")

    return bytes_


def recv_line(sock: socket.socket) -> bytes:
    line = []
    while True:
        c = recv(sock, 1)
        line.append(c)
        if c == b"\n":
            break
    return b"".join(line)


def send(sock: socket.socket, data: Union[bytes, str]) -> int:
    if isinstance(data, str):
        data = data.encode("utf-8")

    if not sock:
        raise WebSocketConnectionClosedException("socket is already closed.")

    def _send():
        try:
            return sock.send(data)
        except SSLWantWriteError:
            pass
        except socket.error as exc:
            error_code = extract_error_code(exc)
            if error_code is None:
                raise
            if error_code not in [errno.EAGAIN, errno.EWOULDBLOCK]:
                raise

        sel = selectors.DefaultSelector()
        sel.register(sock, selectors.EVENT_WRITE)

        w = sel.select(sock.gettimeout())
        sel.close()

        if w:
            return sock.send(data)

    try:
        if sock.gettimeout() == 0:
            return sock.send(data)
        else:
            return _send()
    except socket.timeout as e:
        message = extract_err_message(e)
        raise WebSocketTimeoutException(message)
    except Exception as e:
        message = extract_err_message(e)
        if isinstance(message, str) and "timed out" in message:
            raise WebSocketTimeoutException(message)
        else:
            raise

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\cffi\cffi_opcode.py ===
from .error import VerificationError

class CffiOp(object):
    def __init__(self, op, arg):
        self.op = op
        self.arg = arg

    def as_c_expr(self):
        if self.op is None:
            assert isinstance(self.arg, str)
            return '(_cffi_opcode_t)(%s)' % (self.arg,)
        classname = CLASS_NAME[self.op]
        return '_CFFI_OP(_CFFI_OP_%s, %s)' % (classname, self.arg)

    def as_python_bytes(self):
        if self.op is None and self.arg.isdigit():
            value = int(self.arg)     # non-negative: '-' not in self.arg
            if value >= 2**31:
                raise OverflowError("cannot emit %r: limited to 2**31-1"
                                    % (self.arg,))
            return format_four_bytes(value)
        if isinstance(self.arg, str):
            raise VerificationError("cannot emit to Python: %r" % (self.arg,))
        return format_four_bytes((self.arg << 8) | self.op)

    def __str__(self):
        classname = CLASS_NAME.get(self.op, self.op)
        return '(%s %s)' % (classname, self.arg)

def format_four_bytes(num):
    return '\\x%02X\\x%02X\\x%02X\\x%02X' % (
        (num >> 24) & 0xFF,
        (num >> 16) & 0xFF,
        (num >>  8) & 0xFF,
        (num      ) & 0xFF)

OP_PRIMITIVE       = 1
OP_POINTER         = 3
OP_ARRAY           = 5
OP_OPEN_ARRAY      = 7
OP_STRUCT_UNION    = 9
OP_ENUM            = 11
OP_FUNCTION        = 13
OP_FUNCTION_END    = 15
OP_NOOP            = 17
OP_BITFIELD        = 19
OP_TYPENAME        = 21
OP_CPYTHON_BLTN_V  = 23   # varargs
OP_CPYTHON_BLTN_N  = 25   # noargs
OP_CPYTHON_BLTN_O  = 27   # O  (i.e. a single arg)
OP_CONSTANT        = 29
OP_CONSTANT_INT    = 31
OP_GLOBAL_VAR      = 33
OP_DLOPEN_FUNC     = 35
OP_DLOPEN_CONST    = 37
OP_GLOBAL_VAR_F    = 39
OP_EXTERN_PYTHON   = 41

PRIM_VOID          = 0
PRIM_BOOL          = 1
PRIM_CHAR          = 2
PRIM_SCHAR         = 3
PRIM_UCHAR         = 4
PRIM_SHORT         = 5
PRIM_USHORT        = 6
PRIM_INT           = 7
PRIM_UINT          = 8
PRIM_LONG          = 9
PRIM_ULONG         = 10
PRIM_LONGLONG      = 11
PRIM_ULONGLONG     = 12
PRIM_FLOAT         = 13
PRIM_DOUBLE        = 14
PRIM_LONGDOUBLE    = 15

PRIM_WCHAR         = 16
PRIM_INT8          = 17
PRIM_UINT8         = 18
PRIM_INT16         = 19
PRIM_UINT16        = 20
PRIM_INT32         = 21
PRIM_UINT32        = 22
PRIM_INT64         = 23
PRIM_UINT64        = 24
PRIM_INTPTR        = 25
PRIM_UINTPTR       = 26
PRIM_PTRDIFF       = 27
PRIM_SIZE          = 28
PRIM_SSIZE         = 29
PRIM_INT_LEAST8    = 30
PRIM_UINT_LEAST8   = 31
PRIM_INT_LEAST16   = 32
PRIM_UINT_LEAST16  = 33
PRIM_INT_LEAST32   = 34
PRIM_UINT_LEAST32  = 35
PRIM_INT_LEAST64   = 36
PRIM_UINT_LEAST64  = 37
PRIM_INT_FAST8     = 38
PRIM_UINT_FAST8    = 39
PRIM_INT_FAST16    = 40
PRIM_UINT_FAST16   = 41
PRIM_INT_FAST32    = 42
PRIM_UINT_FAST32   = 43
PRIM_INT_FAST64    = 44
PRIM_UINT_FAST64   = 45
PRIM_INTMAX        = 46
PRIM_UINTMAX       = 47
PRIM_FLOATCOMPLEX  = 48
PRIM_DOUBLECOMPLEX = 49
PRIM_CHAR16        = 50
PRIM_CHAR32        = 51

_NUM_PRIM          = 52
_UNKNOWN_PRIM          = -1
_UNKNOWN_FLOAT_PRIM    = -2
_UNKNOWN_LONG_DOUBLE   = -3

_IO_FILE_STRUCT        = -1

PRIMITIVE_TO_INDEX = {
    'char':               PRIM_CHAR,
    'short':              PRIM_SHORT,
    'int':                PRIM_INT,
    'long':               PRIM_LONG,
    'long long':          PRIM_LONGLONG,
    'signed char':        PRIM_SCHAR,
    'unsigned char':      PRIM_UCHAR,
    'unsigned short':     PRIM_USHORT,
    'unsigned int':       PRIM_UINT,
    'unsigned long':      PRIM_ULONG,
    'unsigned long long': PRIM_ULONGLONG,
    'float':              PRIM_FLOAT,
    'double':             PRIM_DOUBLE,
    'long double':        PRIM_LONGDOUBLE,
    '_cffi_float_complex_t': PRIM_FLOATCOMPLEX,
    '_cffi_double_complex_t': PRIM_DOUBLECOMPLEX,
    '_Bool':              PRIM_BOOL,
    'wchar_t':            PRIM_WCHAR,
    'char16_t':           PRIM_CHAR16,
    'char32_t':           PRIM_CHAR32,
    'int8_t':             PRIM_INT8,
    'uint8_t':            PRIM_UINT8,
    'int16_t':            PRIM_INT16,
    'uint16_t':           PRIM_UINT16,
    'int32_t':            PRIM_INT32,
    'uint32_t':           PRIM_UINT32,
    'int64_t':            PRIM_INT64,
    'uint64_t':           PRIM_UINT64,
    'intptr_t':           PRIM_INTPTR,
    'uintptr_t':          PRIM_UINTPTR,
    'ptrdiff_t':          PRIM_PTRDIFF,
    'size_t':             PRIM_SIZE,
    'ssize_t':            PRIM_SSIZE,
    'int_least8_t':       PRIM_INT_LEAST8,
    'uint_least8_t':      PRIM_UINT_LEAST8,
    'int_least16_t':      PRIM_INT_LEAST16,
    'uint_least16_t':     PRIM_UINT_LEAST16,
    'int_least32_t':      PRIM_INT_LEAST32,
    'uint_least32_t':     PRIM_UINT_LEAST32,
    'int_least64_t':      PRIM_INT_LEAST64,
    'uint_least64_t':     PRIM_UINT_LEAST64,
    'int_fast8_t':        PRIM_INT_FAST8,
    'uint_fast8_t':       PRIM_UINT_FAST8,
    'int_fast16_t':       PRIM_INT_FAST16,
    'uint_fast16_t':      PRIM_UINT_FAST16,
    'int_fast32_t':       PRIM_INT_FAST32,
    'uint_fast32_t':      PRIM_UINT_FAST32,
    'int_fast64_t':       PRIM_INT_FAST64,
    'uint_fast64_t':      PRIM_UINT_FAST64,
    'intmax_t':           PRIM_INTMAX,
    'uintmax_t':          PRIM_UINTMAX,
    }

F_UNION         = 0x01
F_CHECK_FIELDS  = 0x02
F_PACKED        = 0x04
F_EXTERNAL      = 0x08
F_OPAQUE        = 0x10

G_FLAGS = dict([('_CFFI_' + _key, globals()[_key])
                for _key in ['F_UNION', 'F_CHECK_FIELDS', 'F_PACKED',
                             'F_EXTERNAL', 'F_OPAQUE']])

CLASS_NAME = {}
for _name, _value in list(globals().items()):
    if _name.startswith('OP_') and isinstance(_value, int):
        CLASS_NAME[_value] = _name[3:]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\mpmath\functions\factorials.py ===
from ..libmp.backend import xrange
from .functions import defun, defun_wrapped

@defun
def gammaprod(ctx, a, b, _infsign=False):
    a = [ctx.convert(x) for x in a]
    b = [ctx.convert(x) for x in b]
    poles_num = []
    poles_den = []
    regular_num = []
    regular_den = []
    for x in a: [regular_num, poles_num][ctx.isnpint(x)].append(x)
    for x in b: [regular_den, poles_den][ctx.isnpint(x)].append(x)
    # One more pole in numerator or denominator gives 0 or inf
    if len(poles_num) < len(poles_den): return ctx.zero
    if len(poles_num) > len(poles_den):
        # Get correct sign of infinity for x+h, h -> 0 from above
        # XXX: hack, this should be done properly
        if _infsign:
            a = [x and x*(1+ctx.eps) or x+ctx.eps for x in poles_num]
            b = [x and x*(1+ctx.eps) or x+ctx.eps for x in poles_den]
            return ctx.sign(ctx.gammaprod(a+regular_num,b+regular_den)) * ctx.inf
        else:
            return ctx.inf
    # All poles cancel
    # lim G(i)/G(j) = (-1)**(i+j) * gamma(1-j) / gamma(1-i)
    p = ctx.one
    orig = ctx.prec
    try:
        ctx.prec = orig + 15
        while poles_num:
            i = poles_num.pop()
            j = poles_den.pop()
            p *= (-1)**(i+j) * ctx.gamma(1-j) / ctx.gamma(1-i)
        for x in regular_num: p *= ctx.gamma(x)
        for x in regular_den: p /= ctx.gamma(x)
    finally:
        ctx.prec = orig
    return +p

@defun
def beta(ctx, x, y):
    x = ctx.convert(x)
    y = ctx.convert(y)
    if ctx.isinf(y):
        x, y = y, x
    if ctx.isinf(x):
        if x == ctx.inf and not ctx._im(y):
            if y == ctx.ninf:
                return ctx.nan
            if y > 0:
                return ctx.zero
            if ctx.isint(y):
                return ctx.nan
            if y < 0:
                return ctx.sign(ctx.gamma(y)) * ctx.inf
        return ctx.nan
    xy = ctx.fadd(x, y, prec=2*ctx.prec)
    return ctx.gammaprod([x, y], [xy])

@defun
def binomial(ctx, n, k):
    n1 = ctx.fadd(n, 1, prec=2*ctx.prec)
    k1 = ctx.fadd(k, 1, prec=2*ctx.prec)
    nk1 = ctx.fsub(n1, k, prec=2*ctx.prec)
    return ctx.gammaprod([n1], [k1, nk1])

@defun
def rf(ctx, x, n):
    xn = ctx.fadd(x, n, prec=2*ctx.prec)
    return ctx.gammaprod([xn], [x])

@defun
def ff(ctx, x, n):
    x1 = ctx.fadd(x, 1, prec=2*ctx.prec)
    xn1 = ctx.fadd(ctx.fsub(x, n, prec=2*ctx.prec), 1, prec=2*ctx.prec)
    return ctx.gammaprod([x1], [xn1])

@defun_wrapped
def fac2(ctx, x):
    if ctx.isinf(x):
        if x == ctx.inf:
            return x
        return ctx.nan
    return 2**(x/2)*(ctx.pi/2)**((ctx.cospi(x)-1)/4)*ctx.gamma(x/2+1)

@defun_wrapped
def barnesg(ctx, z):
    if ctx.isinf(z):
        if z == ctx.inf:
            return z
        return ctx.nan
    if ctx.isnan(z):
        return z
    if (not ctx._im(z)) and ctx._re(z) <= 0 and ctx.isint(ctx._re(z)):
        return z*0
    # Account for size (would not be needed if computing log(G))
    if abs(z) > 5:
        ctx.dps += 2*ctx.log(abs(z),2)
    # Reflection formula
    if ctx.re(z) < -ctx.dps:
        w = 1-z
        pi2 = 2*ctx.pi
        u = ctx.expjpi(2*w)
        v = ctx.j*ctx.pi/12 - ctx.j*ctx.pi*w**2/2 + w*ctx.ln(1-u) - \
            ctx.j*ctx.polylog(2, u)/pi2
        v = ctx.barnesg(2-z)*ctx.exp(v)/pi2**w
        if ctx._is_real_type(z):
            v = ctx._re(v)
        return v
    # Estimate terms for asymptotic expansion
    # TODO: fixme, obviously
    N = ctx.dps // 2 + 5
    G = 1
    while abs(z) < N or ctx.re(z) < 1:
        G /= ctx.gamma(z)
        z += 1
    z -= 1
    s = ctx.mpf(1)/12
    s -= ctx.log(ctx.glaisher)
    s += z*ctx.log(2*ctx.pi)/2
    s += (z**2/2-ctx.mpf(1)/12)*ctx.log(z)
    s -= 3*z**2/4
    z2k = z2 = z**2
    for k in xrange(1, N+1):
        t = ctx.bernoulli(2*k+2) / (4*k*(k+1)*z2k)
        if abs(t) < ctx.eps:
            #print k, N      # check how many terms were needed
            break
        z2k *= z2
        s += t
    #if k == N:
    #    print "warning: series for barnesg failed to converge", ctx.dps
    return G*ctx.exp(s)

@defun
def superfac(ctx, z):
    return ctx.barnesg(z+2)

@defun_wrapped
def hyperfac(ctx, z):
    # XXX: estimate needed extra bits accurately
    if z == ctx.inf:
        return z
    if abs(z) > 5:
        extra = 4*int(ctx.log(abs(z),2))
    else:
        extra = 0
    ctx.prec += extra
    if not ctx._im(z) and ctx._re(z) < 0 and ctx.isint(ctx._re(z)):
        n = int(ctx.re(z))
        h = ctx.hyperfac(-n-1)
        if ((n+1)//2) & 1:
            h = -h
        if ctx._is_complex_type(z):
            return h + 0j
        return h
    zp1 = z+1
    # Wrong branch cut
    #v = ctx.gamma(zp1)**z
    #ctx.prec -= extra
    #return v / ctx.barnesg(zp1)
    v = ctx.exp(z*ctx.loggamma(zp1))
    ctx.prec -= extra
    return v / ctx.barnesg(zp1)

'''
@defun
def psi0(ctx, z):
    """Shortcut for psi(0,z) (the digamma function)"""
    return ctx.psi(0, z)

@defun
def psi1(ctx, z):
    """Shortcut for psi(1,z) (the trigamma function)"""
    return ctx.psi(1, z)

@defun
def psi2(ctx, z):
    """Shortcut for psi(2,z) (the tetragamma function)"""
    return ctx.psi(2, z)

@defun
def psi3(ctx, z):
    """Shortcut for psi(3,z) (the pentagamma function)"""
    return ctx.psi(3, z)
'''

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\polynomial\__init__.py ===
"""
A sub-package for efficiently dealing with polynomials.

Within the documentation for this sub-package, a "finite power series,"
i.e., a polynomial (also referred to simply as a "series") is represented
by a 1-D numpy array of the polynomial's coefficients, ordered from lowest
order term to highest.  For example, array([1,2,3]) represents
``P_0 + 2*P_1 + 3*P_2``, where P_n is the n-th order basis polynomial
applicable to the specific module in question, e.g., `polynomial` (which
"wraps" the "standard" basis) or `chebyshev`.  For optimal performance,
all operations on polynomials, including evaluation at an argument, are
implemented as operations on the coefficients.  Additional (module-specific)
information can be found in the docstring for the module of interest.

This package provides *convenience classes* for each of six different kinds
of polynomials:

========================    ================
**Name**                    **Provides**
========================    ================
`~polynomial.Polynomial`    Power series
`~chebyshev.Chebyshev`      Chebyshev series
`~legendre.Legendre`        Legendre series
`~laguerre.Laguerre`        Laguerre series
`~hermite.Hermite`          Hermite series
`~hermite_e.HermiteE`       HermiteE series
========================    ================

These *convenience classes* provide a consistent interface for creating,
manipulating, and fitting data with polynomials of different bases.
The convenience classes are the preferred interface for the `~numpy.polynomial`
package, and are available from the ``numpy.polynomial`` namespace.
This eliminates the need to navigate to the corresponding submodules, e.g.
``np.polynomial.Polynomial`` or ``np.polynomial.Chebyshev`` instead of
``np.polynomial.polynomial.Polynomial`` or
``np.polynomial.chebyshev.Chebyshev``, respectively.
The classes provide a more consistent and concise interface than the
type-specific functions defined in the submodules for each type of polynomial.
For example, to fit a Chebyshev polynomial with degree ``1`` to data given
by arrays ``xdata`` and ``ydata``, the
`~chebyshev.Chebyshev.fit` class method::

    >>> from numpy.polynomial import Chebyshev
    >>> xdata = [1, 2, 3, 4]
    >>> ydata = [1, 4, 9, 16]
    >>> c = Chebyshev.fit(xdata, ydata, deg=1)

is preferred over the `chebyshev.chebfit` function from the
``np.polynomial.chebyshev`` module::

    >>> from numpy.polynomial.chebyshev import chebfit
    >>> c = chebfit(xdata, ydata, deg=1)

See :doc:`routines.polynomials.classes` for more details.

Convenience Classes
===================

The following lists the various constants and methods common to all of
the classes representing the various kinds of polynomials. In the following,
the term ``Poly`` represents any one of the convenience classes (e.g.
`~polynomial.Polynomial`, `~chebyshev.Chebyshev`, `~hermite.Hermite`, etc.)
while the lowercase ``p`` represents an **instance** of a polynomial class.

Constants
---------

- ``Poly.domain``     -- Default domain
- ``Poly.window``     -- Default window
- ``Poly.basis_name`` -- String used to represent the basis
- ``Poly.maxpower``   -- Maximum value ``n`` such that ``p**n`` is allowed

Creation
--------

Methods for creating polynomial instances.

- ``Poly.basis(degree)``    -- Basis polynomial of given degree
- ``Poly.identity()``       -- ``p`` where ``p(x) = x`` for all ``x``
- ``Poly.fit(x, y, deg)``   -- ``p`` of degree ``deg`` with coefficients
  determined by the least-squares fit to the data ``x``, ``y``
- ``Poly.fromroots(roots)`` -- ``p`` with specified roots
- ``p.copy()``              -- Create a copy of ``p``

Conversion
----------

Methods for converting a polynomial instance of one kind to another.

- ``p.cast(Poly)``    -- Convert ``p`` to instance of kind ``Poly``
- ``p.convert(Poly)`` -- Convert ``p`` to instance of kind ``Poly`` or map
  between ``domain`` and ``window``

Calculus
--------
- ``p.deriv()`` -- Take the derivative of ``p``
- ``p.integ()`` -- Integrate ``p``

Validation
----------
- ``Poly.has_samecoef(p1, p2)``   -- Check if coefficients match
- ``Poly.has_samedomain(p1, p2)`` -- Check if domains match
- ``Poly.has_sametype(p1, p2)``   -- Check if types match
- ``Poly.has_samewindow(p1, p2)`` -- Check if windows match

Misc
----
- ``p.linspace()`` -- Return ``x, p(x)`` at equally-spaced points in ``domain``
- ``p.mapparms()`` -- Return the parameters for the linear mapping between
  ``domain`` and ``window``.
- ``p.roots()``    -- Return the roots of ``p``.
- ``p.trim()``     -- Remove trailing coefficients.
- ``p.cutdeg(degree)`` -- Truncate ``p`` to given degree
- ``p.truncate(size)`` -- Truncate ``p`` to given size

"""
from .chebyshev import Chebyshev
from .hermite import Hermite
from .hermite_e import HermiteE
from .laguerre import Laguerre
from .legendre import Legendre
from .polynomial import Polynomial

__all__ = [  # noqa: F822
    "set_default_printstyle",
    "polynomial", "Polynomial",
    "chebyshev", "Chebyshev",
    "legendre", "Legendre",
    "hermite", "Hermite",
    "hermite_e", "HermiteE",
    "laguerre", "Laguerre",
]


def set_default_printstyle(style):
    """
    Set the default format for the string representation of polynomials.

    Values for ``style`` must be valid inputs to ``__format__``, i.e. 'ascii'
    or 'unicode'.

    Parameters
    ----------
    style : str
        Format string for default printing style. Must be either 'ascii' or
        'unicode'.

    Notes
    -----
    The default format depends on the platform: 'unicode' is used on
    Unix-based systems and 'ascii' on Windows. This determination is based on
    default font support for the unicode superscript and subscript ranges.

    Examples
    --------
    >>> p = np.polynomial.Polynomial([1, 2, 3])
    >>> c = np.polynomial.Chebyshev([1, 2, 3])
    >>> np.polynomial.set_default_printstyle('unicode')
    >>> print(p)
    1.0 + 2.0·x + 3.0·x²
    >>> print(c)
    1.0 + 2.0·T₁(x) + 3.0·T₂(x)
    >>> np.polynomial.set_default_printstyle('ascii')
    >>> print(p)
    1.0 + 2.0 x + 3.0 x**2
    >>> print(c)
    1.0 + 2.0 T_1(x) + 3.0 T_2(x)
    >>> # Formatting supersedes all class/package-level defaults
    >>> print(f"{p:unicode}")
    1.0 + 2.0·x + 3.0·x²
    """
    if style not in ('unicode', 'ascii'):
        raise ValueError(
            f"Unsupported format string '{style}'. Valid options are 'ascii' "
            f"and 'unicode'"
        )
    _use_unicode = True
    if style == 'ascii':
        _use_unicode = False
    from ._polybase import ABCPolyBase
    ABCPolyBase._use_unicode = _use_unicode


from numpy._pytesttester import PytestTester

test = PytestTester(__name__)
del PytestTester

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\columns.py ===
from collections import defaultdict
from itertools import chain
from operator import itemgetter
from typing import Dict, Iterable, List, Optional, Tuple

from .align import Align, AlignMethod
from .console import Console, ConsoleOptions, RenderableType, RenderResult
from .constrain import Constrain
from .measure import Measurement
from .padding import Padding, PaddingDimensions
from .table import Table
from .text import TextType
from .jupyter import JupyterMixin


class Columns(JupyterMixin):
    """Display renderables in neat columns.

    Args:
        renderables (Iterable[RenderableType]): Any number of Rich renderables (including str).
        width (int, optional): The desired width of the columns, or None to auto detect. Defaults to None.
        padding (PaddingDimensions, optional): Optional padding around cells. Defaults to (0, 1).
        expand (bool, optional): Expand columns to full width. Defaults to False.
        equal (bool, optional): Arrange in to equal sized columns. Defaults to False.
        column_first (bool, optional): Align items from top to bottom (rather than left to right). Defaults to False.
        right_to_left (bool, optional): Start column from right hand side. Defaults to False.
        align (str, optional): Align value ("left", "right", or "center") or None for default. Defaults to None.
        title (TextType, optional): Optional title for Columns.
    """

    def __init__(
        self,
        renderables: Optional[Iterable[RenderableType]] = None,
        padding: PaddingDimensions = (0, 1),
        *,
        width: Optional[int] = None,
        expand: bool = False,
        equal: bool = False,
        column_first: bool = False,
        right_to_left: bool = False,
        align: Optional[AlignMethod] = None,
        title: Optional[TextType] = None,
    ) -> None:
        self.renderables = list(renderables or [])
        self.width = width
        self.padding = padding
        self.expand = expand
        self.equal = equal
        self.column_first = column_first
        self.right_to_left = right_to_left
        self.align: Optional[AlignMethod] = align
        self.title = title

    def add_renderable(self, renderable: RenderableType) -> None:
        """Add a renderable to the columns.

        Args:
            renderable (RenderableType): Any renderable object.
        """
        self.renderables.append(renderable)

    def __rich_console__(
        self, console: Console, options: ConsoleOptions
    ) -> RenderResult:
        render_str = console.render_str
        renderables = [
            render_str(renderable) if isinstance(renderable, str) else renderable
            for renderable in self.renderables
        ]
        if not renderables:
            return
        _top, right, _bottom, left = Padding.unpack(self.padding)
        width_padding = max(left, right)
        max_width = options.max_width
        widths: Dict[int, int] = defaultdict(int)
        column_count = len(renderables)

        get_measurement = Measurement.get
        renderable_widths = [
            get_measurement(console, options, renderable).maximum
            for renderable in renderables
        ]
        if self.equal:
            renderable_widths = [max(renderable_widths)] * len(renderable_widths)

        def iter_renderables(
            column_count: int,
        ) -> Iterable[Tuple[int, Optional[RenderableType]]]:
            item_count = len(renderables)
            if self.column_first:
                width_renderables = list(zip(renderable_widths, renderables))

                column_lengths: List[int] = [item_count // column_count] * column_count
                for col_no in range(item_count % column_count):
                    column_lengths[col_no] += 1

                row_count = (item_count + column_count - 1) // column_count
                cells = [[-1] * column_count for _ in range(row_count)]
                row = col = 0
                for index in range(item_count):
                    cells[row][col] = index
                    column_lengths[col] -= 1
                    if column_lengths[col]:
                        row += 1
                    else:
                        col += 1
                        row = 0
                for index in chain.from_iterable(cells):
                    if index == -1:
                        break
                    yield width_renderables[index]
            else:
                yield from zip(renderable_widths, renderables)
            # Pad odd elements with spaces
            if item_count % column_count:
                for _ in range(column_count - (item_count % column_count)):
                    yield 0, None

        table = Table.grid(padding=self.padding, collapse_padding=True, pad_edge=False)
        table.expand = self.expand
        table.title = self.title

        if self.width is not None:
            column_count = (max_width) // (self.width + width_padding)
            for _ in range(column_count):
                table.add_column(width=self.width)
        else:
            while column_count > 1:
                widths.clear()
                column_no = 0
                for renderable_width, _ in iter_renderables(column_count):
                    widths[column_no] = max(widths[column_no], renderable_width)
                    total_width = sum(widths.values()) + width_padding * (
                        len(widths) - 1
                    )
                    if total_width > max_width:
                        column_count = len(widths) - 1
                        break
                    else:
                        column_no = (column_no + 1) % column_count
                else:
                    break

        get_renderable = itemgetter(1)
        _renderables = [
            get_renderable(_renderable)
            for _renderable in iter_renderables(column_count)
        ]
        if self.equal:
            _renderables = [
                None
                if renderable is None
                else Constrain(renderable, renderable_widths[0])
                for renderable in _renderables
            ]
        if self.align:
            align = self.align
            _Align = Align
            _renderables = [
                None if renderable is None else _Align(renderable, align)
                for renderable in _renderables
            ]

        right_to_left = self.right_to_left
        add_row = table.add_row
        for start in range(0, len(_renderables), column_count):
            row = _renderables[start : start + column_count]
            if right_to_left:
                row = row[::-1]
            add_row(*row)
        yield table


if __name__ == "__main__":  # pragma: no cover
    import os

    console = Console()

    files = [f"{i} {s}" for i, s in enumerate(sorted(os.listdir()))]
    columns = Columns(files, padding=(0, 1), expand=False, equal=False)
    console.print(columns)
    console.rule()
    columns.column_first = True
    console.print(columns)
    columns.right_to_left = True
    console.rule()
    console.print(columns)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\codegen\approximations.py ===
import math
from sympy.sets.sets import Interval
from sympy.calculus.singularities import is_increasing, is_decreasing
from sympy.codegen.rewriting import Optimization
from sympy.core.function import UndefinedFunction

"""
This module collects classes useful for approximate rewriting of expressions.
This can be beneficial when generating numeric code for which performance is
of greater importance than precision (e.g. for preconditioners used in iterative
methods).
"""

class SumApprox(Optimization):
    """
    Approximates sum by neglecting small terms.

    Explanation
    ===========

    If terms are expressions which can be determined to be monotonic, then
    bounds for those expressions are added.

    Parameters
    ==========

    bounds : dict
        Mapping expressions to length 2 tuple of bounds (low, high).
    reltol : number
        Threshold for when to ignore a term. Taken relative to the largest
        lower bound among bounds.

    Examples
    ========

    >>> from sympy import exp
    >>> from sympy.abc import x, y, z
    >>> from sympy.codegen.rewriting import optimize
    >>> from sympy.codegen.approximations import SumApprox
    >>> bounds = {x: (-1, 1), y: (1000, 2000), z: (-10, 3)}
    >>> sum_approx3 = SumApprox(bounds, reltol=1e-3)
    >>> sum_approx2 = SumApprox(bounds, reltol=1e-2)
    >>> sum_approx1 = SumApprox(bounds, reltol=1e-1)
    >>> expr = 3*(x + y + exp(z))
    >>> optimize(expr, [sum_approx3])
    3*(x + y + exp(z))
    >>> optimize(expr, [sum_approx2])
    3*y + 3*exp(z)
    >>> optimize(expr, [sum_approx1])
    3*y

    """

    def __init__(self, bounds, reltol, **kwargs):
        super().__init__(**kwargs)
        self.bounds = bounds
        self.reltol = reltol

    def __call__(self, expr):
        return expr.factor().replace(self.query, lambda arg: self.value(arg))

    def query(self, expr):
        return expr.is_Add

    def value(self, add):
        for term in add.args:
            if term.is_number or term in self.bounds or len(term.free_symbols) != 1:
                continue
            fs, = term.free_symbols
            if fs not in self.bounds:
                continue
            intrvl = Interval(*self.bounds[fs])
            if is_increasing(term, intrvl, fs):
                self.bounds[term] = (
                    term.subs({fs: self.bounds[fs][0]}),
                    term.subs({fs: self.bounds[fs][1]})
                )
            elif is_decreasing(term, intrvl, fs):
                self.bounds[term] = (
                    term.subs({fs: self.bounds[fs][1]}),
                    term.subs({fs: self.bounds[fs][0]})
                )
            else:
                return add

        if all(term.is_number or term in self.bounds for term in add.args):
            bounds = [(term, term) if term.is_number else self.bounds[term] for term in add.args]
            largest_abs_guarantee = 0
            for lo, hi in bounds:
                if lo <= 0 <= hi:
                    continue
                largest_abs_guarantee = max(largest_abs_guarantee,
                                            min(abs(lo), abs(hi)))
            new_terms = []
            for term, (lo, hi) in zip(add.args, bounds):
                if max(abs(lo), abs(hi)) >= largest_abs_guarantee*self.reltol:
                    new_terms.append(term)
            return add.func(*new_terms)
        else:
            return add


class SeriesApprox(Optimization):
    """ Approximates functions by expanding them as a series.

    Parameters
    ==========

    bounds : dict
        Mapping expressions to length 2 tuple of bounds (low, high).
    reltol : number
        Threshold for when to ignore a term. Taken relative to the largest
        lower bound among bounds.
    max_order : int
        Largest order to include in series expansion
    n_point_checks : int (even)
        The validity of an expansion (with respect to reltol) is checked at
        discrete points (linearly spaced over the bounds of the variable). The
        number of points used in this numerical check is given by this number.

    Examples
    ========

    >>> from sympy import sin, pi
    >>> from sympy.abc import x, y
    >>> from sympy.codegen.rewriting import optimize
    >>> from sympy.codegen.approximations import SeriesApprox
    >>> bounds = {x: (-.1, .1), y: (pi-1, pi+1)}
    >>> series_approx2 = SeriesApprox(bounds, reltol=1e-2)
    >>> series_approx3 = SeriesApprox(bounds, reltol=1e-3)
    >>> series_approx8 = SeriesApprox(bounds, reltol=1e-8)
    >>> expr = sin(x)*sin(y)
    >>> optimize(expr, [series_approx2])
    x*(-y + (y - pi)**3/6 + pi)
    >>> optimize(expr, [series_approx3])
    (-x**3/6 + x)*sin(y)
    >>> optimize(expr, [series_approx8])
    sin(x)*sin(y)

    """
    def __init__(self, bounds, reltol, max_order=4, n_point_checks=4, **kwargs):
        super().__init__(**kwargs)
        self.bounds = bounds
        self.reltol = reltol
        self.max_order = max_order
        if n_point_checks % 2 == 1:
            raise ValueError("Checking the solution at expansion point is not helpful")
        self.n_point_checks = n_point_checks
        self._prec = math.ceil(-math.log10(self.reltol))

    def __call__(self, expr):
        return expr.factor().replace(self.query, lambda arg: self.value(arg))

    def query(self, expr):
        return (expr.is_Function and not isinstance(expr, UndefinedFunction)
                and len(expr.args) == 1)

    def value(self, fexpr):
        free_symbols = fexpr.free_symbols
        if len(free_symbols) != 1:
            return fexpr
        symb, = free_symbols
        if symb not in self.bounds:
            return fexpr
        lo, hi = self.bounds[symb]
        x0 = (lo + hi)/2
        cheapest = None
        for n in range(self.max_order+1, 0, -1):
            fseri = fexpr.series(symb, x0=x0, n=n).removeO()
            n_ok = True
            for idx in range(self.n_point_checks):
                x = lo + idx*(hi - lo)/(self.n_point_checks - 1)
                val = fseri.xreplace({symb: x})
                ref = fexpr.xreplace({symb: x})
                if abs((1 - val/ref).evalf(self._prec)) > self.reltol:
                    n_ok = False
                    break

            if n_ok:
                cheapest = fseri
            else:
                break

        if cheapest is None:
            return fexpr
        else:
            return cheapest

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\external\importtools.py ===
"""Tools to assist importing optional external modules."""

import sys
import re

# Override these in the module to change the default warning behavior.
# For example, you might set both to False before running the tests so that
# warnings are not printed to the console, or set both to True for debugging.

WARN_NOT_INSTALLED = None  # Default is False
WARN_OLD_VERSION = None  # Default is True


def __sympy_debug():
    # helper function from sympy/__init__.py
    # We don't just import SYMPY_DEBUG from that file because we don't want to
    # import all of SymPy just to use this module.
    import os
    debug_str = os.getenv('SYMPY_DEBUG', 'False')
    if debug_str in ('True', 'False'):
        return eval(debug_str)
    else:
        raise RuntimeError("unrecognized value for SYMPY_DEBUG: %s" %
                           debug_str)

if __sympy_debug():
    WARN_OLD_VERSION = True
    WARN_NOT_INSTALLED = True


_component_re = re.compile(r'(\d+ | [a-z]+ | \.)', re.VERBOSE)

def version_tuple(vstring):
    # Parse a version string to a tuple e.g. '1.2' -> (1, 2)
    # Simplified from distutils.version.LooseVersion which was deprecated in
    # Python 3.10.
    components = []
    for x in _component_re.split(vstring):
        if x and x != '.':
            try:
                x = int(x)
            except ValueError:
                pass
            components.append(x)
    return tuple(components)


def import_module(module, min_module_version=None, min_python_version=None,
        warn_not_installed=None, warn_old_version=None,
        module_version_attr='__version__', module_version_attr_call_args=None,
        import_kwargs={}, catch=()):
    """
    Import and return a module if it is installed.

    If the module is not installed, it returns None.

    A minimum version for the module can be given as the keyword argument
    min_module_version.  This should be comparable against the module version.
    By default, module.__version__ is used to get the module version.  To
    override this, set the module_version_attr keyword argument.  If the
    attribute of the module to get the version should be called (e.g.,
    module.version()), then set module_version_attr_call_args to the args such
    that module.module_version_attr(*module_version_attr_call_args) returns the
    module's version.

    If the module version is less than min_module_version using the Python <
    comparison, None will be returned, even if the module is installed. You can
    use this to keep from importing an incompatible older version of a module.

    You can also specify a minimum Python version by using the
    min_python_version keyword argument.  This should be comparable against
    sys.version_info.

    If the keyword argument warn_not_installed is set to True, the function will
    emit a UserWarning when the module is not installed.

    If the keyword argument warn_old_version is set to True, the function will
    emit a UserWarning when the library is installed, but cannot be imported
    because of the min_module_version or min_python_version options.

    Note that because of the way warnings are handled, a warning will be
    emitted for each module only once.  You can change the default warning
    behavior by overriding the values of WARN_NOT_INSTALLED and WARN_OLD_VERSION
    in sympy.external.importtools.  By default, WARN_NOT_INSTALLED is False and
    WARN_OLD_VERSION is True.

    This function uses __import__() to import the module.  To pass additional
    options to __import__(), use the import_kwargs keyword argument.  For
    example, to import a submodule A.B, you must pass a nonempty fromlist option
    to __import__.  See the docstring of __import__().

    This catches ImportError to determine if the module is not installed.  To
    catch additional errors, pass them as a tuple to the catch keyword
    argument.

    Examples
    ========

    >>> from sympy.external import import_module

    >>> numpy = import_module('numpy')

    >>> numpy = import_module('numpy', min_python_version=(2, 7),
    ... warn_old_version=False)

    >>> numpy = import_module('numpy', min_module_version='1.5',
    ... warn_old_version=False) # numpy.__version__ is a string

    >>> # gmpy does not have __version__, but it does have gmpy.version()

    >>> gmpy = import_module('gmpy', min_module_version='1.14',
    ... module_version_attr='version', module_version_attr_call_args=(),
    ... warn_old_version=False)

    >>> # To import a submodule, you must pass a nonempty fromlist to
    >>> # __import__().  The values do not matter.
    >>> p3 = import_module('mpl_toolkits.mplot3d',
    ... import_kwargs={'fromlist':['something']})

    >>> # matplotlib.pyplot can raise RuntimeError when the display cannot be opened
    >>> matplotlib = import_module('matplotlib',
    ... import_kwargs={'fromlist':['pyplot']}, catch=(RuntimeError,))

    """
    # keyword argument overrides default, and global variable overrides
    # keyword argument.
    warn_old_version = (WARN_OLD_VERSION if WARN_OLD_VERSION is not None
        else warn_old_version or True)
    warn_not_installed = (WARN_NOT_INSTALLED if WARN_NOT_INSTALLED is not None
        else warn_not_installed or False)

    import warnings

    # Check Python first so we don't waste time importing a module we can't use
    if min_python_version:
        if sys.version_info < min_python_version:
            if warn_old_version:
                warnings.warn("Python version is too old to use %s "
                    "(%s or newer required)" % (
                        module, '.'.join(map(str, min_python_version))),
                    UserWarning, stacklevel=2)
            return

    try:
        mod = __import__(module, **import_kwargs)

        ## there's something funny about imports with matplotlib and py3k. doing
        ##    from matplotlib import collections
        ## gives python's stdlib collections module. explicitly re-importing
        ## the module fixes this.
        from_list = import_kwargs.get('fromlist', ())
        for submod in from_list:
            if submod == 'collections' and mod.__name__ == 'matplotlib':
                __import__(module + '.' + submod)
    except ImportError:
        if warn_not_installed:
            warnings.warn("%s module is not installed" % module, UserWarning,
                    stacklevel=2)
        return
    except catch as e:
        if warn_not_installed:
            warnings.warn(
                "%s module could not be used (%s)" % (module, repr(e)),
                stacklevel=2)
        return

    if min_module_version:
        modversion = getattr(mod, module_version_attr)
        if module_version_attr_call_args is not None:
            modversion = modversion(*module_version_attr_call_args)
        if version_tuple(modversion) < version_tuple(min_module_version):
            if warn_old_version:
                # Attempt to create a pretty string version of the version
                if isinstance(min_module_version, str):
                    verstr = min_module_version
                elif isinstance(min_module_version, (tuple, list)):
                    verstr = '.'.join(map(str, min_module_version))
                else:
                    # Either don't know what this is.  Hopefully
                    # it's something that has a nice str version, like an int.
                    verstr = str(min_module_version)
                warnings.warn("%s version is too old to use "
                    "(%s or newer required)" % (module, verstr),
                    UserWarning, stacklevel=2)
            return

    return mod