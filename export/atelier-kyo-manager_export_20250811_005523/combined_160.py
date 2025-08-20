
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\models\stable_diffusion\ort_optimizer.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

"""
ONNX Model Optimizer for Stable Diffusion
"""

import gc
import logging
import os
import shutil
import tempfile
from pathlib import Path

import onnx
from packaging import version

from onnxruntime.transformers.fusion_options import FusionOptions
from onnxruntime.transformers.onnx_model_clip import ClipOnnxModel
from onnxruntime.transformers.onnx_model_unet import UnetOnnxModel
from onnxruntime.transformers.onnx_model_vae import VaeOnnxModel
from onnxruntime.transformers.optimizer import optimize_by_onnxruntime, optimize_model

logger = logging.getLogger(__name__)


class OrtStableDiffusionOptimizer:
    def __init__(self, model_type: str):
        assert model_type in ["vae", "unet", "clip"]
        self.model_type = model_type
        self.model_type_class_mapping = {
            "unet": UnetOnnxModel,
            "vae": VaeOnnxModel,
            "clip": ClipOnnxModel,
        }

    def _optimize_by_ort(self, onnx_model, use_external_data_format, tmp_dir):
        # Save to a temporary file so that we can load it with Onnx Runtime.
        logger.info("Saving a temporary model to run OnnxRuntime graph optimizations...")
        tmp_model_path = Path(tmp_dir) / "model.onnx"
        onnx_model.save_model_to_file(str(tmp_model_path), use_external_data_format=use_external_data_format)

        del onnx_model
        gc.collect()

        ort_optimized_model_path = Path(tmp_dir) / "optimized.onnx"
        optimize_by_onnxruntime(
            str(tmp_model_path),
            use_gpu=True,
            optimized_model_path=str(ort_optimized_model_path),
            save_as_external_data=use_external_data_format,
            external_data_filename="optimized.onnx_data",
        )
        model = onnx.load(str(ort_optimized_model_path), load_external_data=True)
        return self.model_type_class_mapping[self.model_type](model)

    def optimize_by_ort(self, onnx_model, use_external_data_format=False, tmp_dir=None):
        # Use this step to see the final graph that executed by Onnx Runtime.
        if tmp_dir is None:
            with tempfile.TemporaryDirectory() as temp_dir:
                return self._optimize_by_ort(onnx_model, use_external_data_format, temp_dir)
        else:
            os.makedirs(tmp_dir, exist_ok=True)
            model = self._optimize_by_ort(onnx_model, use_external_data_format, tmp_dir)
            shutil.rmtree(tmp_dir)
            return model

    def optimize(
        self,
        input_fp32_onnx_path,
        optimized_onnx_path,
        float16=True,
        keep_io_types=False,
        fp32_op_list=None,
        keep_outputs=None,
        optimize_by_ort=True,
        optimize_by_fusion=True,
        final_target_float16=True,
        tmp_dir=None,
    ):
        """Optimize onnx model using ONNX Runtime transformers optimizer"""
        logger.info(f"Optimize {input_fp32_onnx_path}...")

        if optimize_by_fusion:
            fusion_options = FusionOptions(self.model_type)

            # It is allowed float16=False and final_target_float16=True, for using fp32 as intermediate optimization step.
            # For rare fp32 use case, we can disable packed kv/qkv since there is no fp32 TRT fused attention kernel.
            if self.model_type in ["unet"] and not final_target_float16:
                fusion_options.enable_packed_kv = False
                fusion_options.enable_packed_qkv = False

            m = optimize_model(
                input_fp32_onnx_path,
                model_type=self.model_type,
                num_heads=0,  # will be deduced from graph
                hidden_size=0,  # will be deduced from graph
                opt_level=0,
                optimization_options=fusion_options,
                use_gpu=True,
            )
        else:
            model = onnx.load_model(input_fp32_onnx_path, load_external_data=True)
            m = self.model_type_class_mapping[self.model_type](model)

        if keep_outputs:
            m.prune_graph(outputs=keep_outputs)

        model_size = m.model.ByteSize()

        # model size might be negative (overflow?) in Windows.
        use_external_data_format = model_size <= 0 or model_size >= onnx.checker.MAXIMUM_PROTOBUF

        # Note that ORT < 1.16 could not save model larger than 2GB.
        # This step is is optional since it has no impact on inference latency.
        # The optimized model is not portable. It could only run in the same execution provider (CUDA EP in this case).
        # When the model has been optimized by onnxruntime, we can disable optimization in SessionOption
        # to save session creation time. Another benefit is to inspect the final graph for developing purpose.
        from onnxruntime import __version__ as ort_version

        if optimize_by_ort and (version.parse(ort_version) >= version.parse("1.16.0") or not use_external_data_format):
            m = self.optimize_by_ort(m, use_external_data_format=use_external_data_format, tmp_dir=tmp_dir)

        if float16:
            logger.info("Convert to float16 ...")
            m.convert_float_to_float16(
                keep_io_types=keep_io_types,
                op_block_list=fp32_op_list,
            )

        m.get_operator_statistics()
        m.get_fused_operator_statistics()
        m.save_model_to_file(optimized_onnx_path, use_external_data_format=use_external_data_format)
        logger.info("%s is optimized: %s", self.model_type, optimized_onnx_path)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\cell\read_only.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import is_date_format
from openpyxl.styles.numbers import BUILTIN_FORMATS, BUILTIN_FORMATS_MAX_SIZE


class ReadOnlyCell:

    __slots__ =  ('parent', 'row', 'column', '_value', 'data_type', '_style_id')

    def __init__(self, sheet, row, column, value, data_type='n', style_id=0):
        self.parent = sheet
        self._value = None
        self.row = row
        self.column = column
        self.data_type = data_type
        self.value = value
        self._style_id = style_id


    def __eq__(self, other):
        for a in self.__slots__:
            if getattr(self, a) != getattr(other, a):
                return
        return True

    def __ne__(self, other):
        return not self.__eq__(other)


    def __repr__(self):
        return "<ReadOnlyCell {0!r}.{1}>".format(self.parent.title, self.coordinate)


    @property
    def coordinate(self):
        column = get_column_letter(self.column)
        return "{1}{0}".format(self.row, column)


    @property
    def coordinate(self):
        return Cell.coordinate.__get__(self)


    @property
    def column_letter(self):
        return Cell.column_letter.__get__(self)


    @property
    def style_array(self):
        return self.parent.parent._cell_styles[self._style_id]


    @property
    def has_style(self):
        return self._style_id != 0


    @property
    def number_format(self):
        _id = self.style_array.numFmtId
        if _id < BUILTIN_FORMATS_MAX_SIZE:
            return BUILTIN_FORMATS.get(_id, "General")
        else:
            return self.parent.parent._number_formats[
                _id - BUILTIN_FORMATS_MAX_SIZE]

    @property
    def font(self):
        _id = self.style_array.fontId
        return self.parent.parent._fonts[_id]

    @property
    def fill(self):
        _id = self.style_array.fillId
        return self.parent.parent._fills[_id]

    @property
    def border(self):
        _id = self.style_array.borderId
        return self.parent.parent._borders[_id]

    @property
    def alignment(self):
        _id = self.style_array.alignmentId
        return self.parent.parent._alignments[_id]

    @property
    def protection(self):
        _id = self.style_array.protectionId
        return self.parent.parent._protections[_id]


    @property
    def is_date(self):
        return Cell.is_date.__get__(self)


    @property
    def internal_value(self):
        return self._value

    @property
    def value(self):
        return self._value

    @value.setter
    def value(self, value):
        if self._value is not None:
            raise AttributeError("Cell is read only")
        self._value = value


class EmptyCell:

    __slots__ = ()

    value = None
    is_date = False
    font = None
    border = None
    fill = None
    number_format = None
    alignment = None
    data_type = 'n'


    def __repr__(self):
        return "<EmptyCell>"

EMPTY_CELL = EmptyCell()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\cell\_writer.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.compat import safe_string
from openpyxl.xml.functions import Element, SubElement, whitespace, XML_NS
from openpyxl import LXML
from openpyxl.utils.datetime import to_excel, to_ISO8601
from datetime import timedelta

from openpyxl.worksheet.formula import DataTableFormula, ArrayFormula
from openpyxl.cell.rich_text import CellRichText

def _set_attributes(cell, styled=None):
    """
    Set coordinate and datatype
    """
    coordinate = cell.coordinate
    attrs = {'r': coordinate}
    if styled:
        attrs['s'] = f"{cell.style_id}"

    if cell.data_type == "s":
        attrs['t'] = "inlineStr"
    elif cell.data_type != 'f':
        attrs['t'] = cell.data_type

    value = cell._value

    if cell.data_type == "d":
        if hasattr(value, "tzinfo") and value.tzinfo is not None:
            raise TypeError("Excel does not support timezones in datetimes. "
                    "The tzinfo in the datetime/time object must be set to None.")

        if cell.parent.parent.iso_dates and not isinstance(value, timedelta):
            value = to_ISO8601(value)
        else:
            attrs['t'] = "n"
            value = to_excel(value, cell.parent.parent.epoch)

    if cell.hyperlink:
        cell.parent._hyperlinks.append(cell.hyperlink)

    return value, attrs


def etree_write_cell(xf, worksheet, cell, styled=None):

    value, attributes = _set_attributes(cell, styled)

    el = Element("c", attributes)
    if value is None or value == "":
        xf.write(el)
        return

    if cell.data_type == 'f':
        attrib = {}

        if isinstance(value, ArrayFormula):
            attrib = dict(value)
            value = value.text

        elif isinstance(value, DataTableFormula):
            attrib = dict(value)
            value = None

        formula = SubElement(el, 'f', attrib)
        if value is not None and not attrib.get('t') == "dataTable":
            formula.text = value[1:]
            value = None

    if cell.data_type == 's':
        if isinstance(value, CellRichText):
            el.append(value.to_tree())
        else:
            inline_string = Element("is")
            text = Element('t')
            text.text = value
            whitespace(text)
            inline_string.append(text)
            el.append(inline_string)

    else:
        cell_content = SubElement(el, 'v')
        if value is not None:
            cell_content.text = safe_string(value)

    xf.write(el)


def lxml_write_cell(xf, worksheet, cell, styled=False):
    value, attributes = _set_attributes(cell, styled)

    if value == '' or value is None:
        with xf.element("c", attributes):
            return

    with xf.element('c', attributes):
        if cell.data_type == 'f':
            attrib = {}

            if isinstance(value, ArrayFormula):
                attrib = dict(value)
                value = value.text

            elif isinstance(value, DataTableFormula):
                attrib = dict(value)
                value = None

            with xf.element('f', attrib):
                if value is not None and not attrib.get('t') == "dataTable":
                    xf.write(value[1:])
                    value = None

        if cell.data_type == 's':
            if isinstance(value, CellRichText):
                el = value.to_tree()
                xf.write(el)
            else:
                with xf.element("is"):
                    if isinstance(value, str):
                        attrs = {}
                        if value != value.strip():
                            attrs["{%s}space" % XML_NS] = "preserve"
                        el = Element("t", attrs) # lxml can't handle xml-ns
                        el.text = value
                        xf.write(el)

        else:
            with xf.element("v"):
                if value is not None:
                    xf.write(safe_string(value))


if LXML:
    write_cell = lxml_write_cell
else:
    write_cell = etree_write_cell

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\descriptors\sequence.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.compat import safe_string
from openpyxl.xml.functions import Element
from openpyxl.utils.indexed_list import IndexedList

from .base import Descriptor, Alias, _convert
from .namespace import namespaced


class Sequence(Descriptor):
    """
    A sequence (list or tuple) that may only contain objects of the declared
    type
    """

    expected_type = type(None)
    seq_types = (list, tuple)
    idx_base = 0
    unique = False
    container = list


    def __set__(self, instance, seq):
        if not isinstance(seq, self.seq_types):
            raise TypeError("Value must be a sequence")
        seq = self.container(_convert(self.expected_type, value) for value in seq)
        if self.unique:
            seq = IndexedList(seq)

        super().__set__(instance, seq)


    def to_tree(self, tagname, obj, namespace=None):
        """
        Convert the sequence represented by the descriptor to an XML element
        """
        for idx, v in enumerate(obj, self.idx_base):
            if hasattr(v, "to_tree"):
                el = v.to_tree(tagname, idx)
            else:
                tagname = namespaced(obj, tagname, namespace)
                el = Element(tagname)
                el.text = safe_string(v)
            yield el


class UniqueSequence(Sequence):
    """
    Use a set to keep values unique
    """
    seq_types = (list, tuple, set)
    container = set


class ValueSequence(Sequence):
    """
    A sequence of primitive types that are stored as a single attribute.
    "val" is the default attribute
    """

    attribute = "val"


    def to_tree(self, tagname, obj, namespace=None):
        tagname = namespaced(self, tagname, namespace)
        for v in obj:
            yield Element(tagname, {self.attribute:safe_string(v)})


    def from_tree(self, node):

        return node.get(self.attribute)


class NestedSequence(Sequence):
    """
    Wrap a sequence in an containing object
    """

    count = False

    def to_tree(self, tagname, obj, namespace=None):
        tagname = namespaced(self, tagname, namespace)
        container = Element(tagname)
        if self.count:
            container.set('count', str(len(obj)))
        for v in obj:
            container.append(v.to_tree())
        return container


    def from_tree(self, node):
        return [self.expected_type.from_tree(el) for el in node]


class MultiSequence(Sequence):
    """
    Sequences can contain objects with different tags
    """

    def __set__(self, instance, seq):
        if not isinstance(seq, (tuple, list)):
            raise ValueError("Value must be a sequence")
        seq = list(seq)
        Descriptor.__set__(self, instance, seq)


    def to_tree(self, tagname, obj, namespace=None):
        """
        Convert the sequence represented by the descriptor to an XML element
        """
        for v in obj:
            el = v.to_tree(namespace=namespace)
            yield el


class MultiSequencePart(Alias):
    """
    Allow a multisequence to be built up from parts

    Excluded from the instance __elements__ or __attrs__ as is effectively an Alias
    """

    def __init__(self, expected_type, store):
        self.expected_type = expected_type
        self.store = store


    def __set__(self, instance, value):
        value = _convert(self.expected_type, value)
        instance.__dict__[self.store].append(value)


    def __get__(self, instance, cls):
        return self

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\interchange\buffer.py ===
from __future__ import annotations

from typing import (
    TYPE_CHECKING,
    Any,
)

from pandas.core.interchange.dataframe_protocol import (
    Buffer,
    DlpackDeviceType,
)

if TYPE_CHECKING:
    import numpy as np
    import pyarrow as pa


class PandasBuffer(Buffer):
    """
    Data in the buffer is guaranteed to be contiguous in memory.
    """

    def __init__(self, x: np.ndarray, allow_copy: bool = True) -> None:
        """
        Handle only regular columns (= numpy arrays) for now.
        """
        if x.strides[0] and not x.strides == (x.dtype.itemsize,):
            # The protocol does not support strided buffers, so a copy is
            # necessary. If that's not allowed, we need to raise an exception.
            if allow_copy:
                x = x.copy()
            else:
                raise RuntimeError(
                    "Exports cannot be zero-copy in the case "
                    "of a non-contiguous buffer"
                )

        # Store the numpy array in which the data resides as a private
        # attribute, so we can use it to retrieve the public attributes
        self._x = x

    @property
    def bufsize(self) -> int:
        """
        Buffer size in bytes.
        """
        return self._x.size * self._x.dtype.itemsize

    @property
    def ptr(self) -> int:
        """
        Pointer to start of the buffer as an integer.
        """
        return self._x.__array_interface__["data"][0]

    def __dlpack__(self) -> Any:
        """
        Represent this structure as DLPack interface.
        """
        return self._x.__dlpack__()

    def __dlpack_device__(self) -> tuple[DlpackDeviceType, int | None]:
        """
        Device type and device ID for where the data in the buffer resides.
        """
        return (DlpackDeviceType.CPU, None)

    def __repr__(self) -> str:
        return (
            "PandasBuffer("
            + str(
                {
                    "bufsize": self.bufsize,
                    "ptr": self.ptr,
                    "device": self.__dlpack_device__()[0].name,
                }
            )
            + ")"
        )


class PandasBufferPyarrow(Buffer):
    """
    Data in the buffer is guaranteed to be contiguous in memory.
    """

    def __init__(
        self,
        buffer: pa.Buffer,
        *,
        length: int,
    ) -> None:
        """
        Handle pyarrow chunked arrays.
        """
        self._buffer = buffer
        self._length = length

    @property
    def bufsize(self) -> int:
        """
        Buffer size in bytes.
        """
        return self._buffer.size

    @property
    def ptr(self) -> int:
        """
        Pointer to start of the buffer as an integer.
        """
        return self._buffer.address

    def __dlpack__(self) -> Any:
        """
        Represent this structure as DLPack interface.
        """
        raise NotImplementedError()

    def __dlpack_device__(self) -> tuple[DlpackDeviceType, int | None]:
        """
        Device type and device ID for where the data in the buffer resides.
        """
        return (DlpackDeviceType.CPU, None)

    def __repr__(self) -> str:
        return (
            "PandasBuffer[pyarrow]("
            + str(
                {
                    "bufsize": self.bufsize,
                    "ptr": self.ptr,
                    "device": "CPU",
                }
            )
            + ")"
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\commands\completion.py ===
import sys
import textwrap
from optparse import Values
from typing import List

from pip._internal.cli.base_command import Command
from pip._internal.cli.status_codes import SUCCESS
from pip._internal.utils.misc import get_prog

BASE_COMPLETION = """
# pip {shell} completion start{script}# pip {shell} completion end
"""

COMPLETION_SCRIPTS = {
    "bash": """
        _pip_completion()
        {{
            COMPREPLY=( $( COMP_WORDS="${{COMP_WORDS[*]}}" \\
                           COMP_CWORD=$COMP_CWORD \\
                           PIP_AUTO_COMPLETE=1 $1 2>/dev/null ) )
        }}
        complete -o default -F _pip_completion {prog}
    """,
    "zsh": """
        #compdef -P pip[0-9.]#
        __pip() {{
          compadd $( COMP_WORDS="$words[*]" \\
                     COMP_CWORD=$((CURRENT-1)) \\
                     PIP_AUTO_COMPLETE=1 $words[1] 2>/dev/null )
        }}
        if [[ $zsh_eval_context[-1] == loadautofunc ]]; then
          # autoload from fpath, call function directly
          __pip "$@"
        else
          # eval/source/. command, register function for later
          compdef __pip -P 'pip[0-9.]#'
        fi
    """,
    "fish": """
        function __fish_complete_pip
            set -lx COMP_WORDS \\
                (commandline --current-process --tokenize --cut-at-cursor) \\
                (commandline --current-token --cut-at-cursor)
            set -lx COMP_CWORD (math (count $COMP_WORDS) - 1)
            set -lx PIP_AUTO_COMPLETE 1
            set -l completions
            if string match -q '2.*' $version
                set completions (eval $COMP_WORDS[1])
            else
                set completions ($COMP_WORDS[1])
            end
            string split \\  -- $completions
        end
        complete -fa "(__fish_complete_pip)" -c {prog}
    """,
    "powershell": """
        if ((Test-Path Function:\\TabExpansion) -and -not `
            (Test-Path Function:\\_pip_completeBackup)) {{
            Rename-Item Function:\\TabExpansion _pip_completeBackup
        }}
        function TabExpansion($line, $lastWord) {{
            $lastBlock = [regex]::Split($line, '[|;]')[-1].TrimStart()
            if ($lastBlock.StartsWith("{prog} ")) {{
                $Env:COMP_WORDS=$lastBlock
                $Env:COMP_CWORD=$lastBlock.Split().Length - 1
                $Env:PIP_AUTO_COMPLETE=1
                (& {prog}).Split()
                Remove-Item Env:COMP_WORDS
                Remove-Item Env:COMP_CWORD
                Remove-Item Env:PIP_AUTO_COMPLETE
            }}
            elseif (Test-Path Function:\\_pip_completeBackup) {{
                # Fall back on existing tab expansion
                _pip_completeBackup $line $lastWord
            }}
        }}
    """,
}


class CompletionCommand(Command):
    """A helper command to be used for command completion."""

    ignore_require_venv = True

    def add_options(self) -> None:
        self.cmd_opts.add_option(
            "--bash",
            "-b",
            action="store_const",
            const="bash",
            dest="shell",
            help="Emit completion code for bash",
        )
        self.cmd_opts.add_option(
            "--zsh",
            "-z",
            action="store_const",
            const="zsh",
            dest="shell",
            help="Emit completion code for zsh",
        )
        self.cmd_opts.add_option(
            "--fish",
            "-f",
            action="store_const",
            const="fish",
            dest="shell",
            help="Emit completion code for fish",
        )
        self.cmd_opts.add_option(
            "--powershell",
            "-p",
            action="store_const",
            const="powershell",
            dest="shell",
            help="Emit completion code for powershell",
        )

        self.parser.insert_option_group(0, self.cmd_opts)

    def run(self, options: Values, args: List[str]) -> int:
        """Prints the completion code of the given shell"""
        shells = COMPLETION_SCRIPTS.keys()
        shell_options = ["--" + shell for shell in sorted(shells)]
        if options.shell in shells:
            script = textwrap.dedent(
                COMPLETION_SCRIPTS.get(options.shell, "").format(prog=get_prog())
            )
            print(BASE_COMPLETION.format(script=script, shell=options.shell))
            return SUCCESS
        else:
            sys.stderr.write(
                "ERROR: You must pass {}\n".format(" or ".join(shell_options))
            )
            return SUCCESS

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\mysql\pymysql.py ===
# dialects/mysql/pymysql.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors


r"""

.. dialect:: mysql+pymysql
    :name: PyMySQL
    :dbapi: pymysql
    :connectstring: mysql+pymysql://<username>:<password>@<host>/<dbname>[?<options>]
    :url: https://pymysql.readthedocs.io/

Unicode
-------

Please see :ref:`mysql_unicode` for current recommendations on unicode
handling.

.. _pymysql_ssl:

SSL Connections
------------------

The PyMySQL DBAPI accepts the same SSL arguments as that of MySQLdb,
described at :ref:`mysqldb_ssl`.   See that section for additional examples.

If the server uses an automatically-generated certificate that is self-signed
or does not match the host name (as seen from the client), it may also be
necessary to indicate ``ssl_check_hostname=false`` in PyMySQL::

    connection_uri = (
        "mysql+pymysql://scott:tiger@192.168.0.134/test"
        "?ssl_ca=/home/gord/client-ssl/ca.pem"
        "&ssl_cert=/home/gord/client-ssl/client-cert.pem"
        "&ssl_key=/home/gord/client-ssl/client-key.pem"
        "&ssl_check_hostname=false"
    )

MySQL-Python Compatibility
--------------------------

The pymysql DBAPI is a pure Python port of the MySQL-python (MySQLdb) driver,
and targets 100% compatibility.   Most behavioral notes for MySQL-python apply
to the pymysql driver as well.

"""  # noqa

from .mysqldb import MySQLDialect_mysqldb
from ...util import langhelpers


class MySQLDialect_pymysql(MySQLDialect_mysqldb):
    driver = "pymysql"
    supports_statement_cache = True

    description_encoding = None

    @langhelpers.memoized_property
    def supports_server_side_cursors(self):
        try:
            cursors = __import__("pymysql.cursors").cursors
            self._sscursor = cursors.SSCursor
            return True
        except (ImportError, AttributeError):
            return False

    @classmethod
    def import_dbapi(cls):
        return __import__("pymysql")

    @langhelpers.memoized_property
    def _send_false_to_ping(self):
        """determine if pymysql has deprecated, changed the default of,
        or removed the 'reconnect' argument of connection.ping().

        See #10492 and
        https://github.com/PyMySQL/mysqlclient/discussions/651#discussioncomment-7308971
        for background.

        """  # noqa: E501

        try:
            Connection = __import__(
                "pymysql.connections"
            ).connections.Connection
        except (ImportError, AttributeError):
            return True
        else:
            insp = langhelpers.get_callable_argspec(Connection.ping)
            try:
                reconnect_arg = insp.args[1]
            except IndexError:
                return False
            else:
                return reconnect_arg == "reconnect" and (
                    not insp.defaults or insp.defaults[0] is not False
                )

    def do_ping(self, dbapi_connection):
        if self._send_false_to_ping:
            dbapi_connection.ping(False)
        else:
            dbapi_connection.ping()

        return True

    def create_connect_args(self, url, _translate_args=None):
        if _translate_args is None:
            _translate_args = dict(username="user")
        return super().create_connect_args(
            url, _translate_args=_translate_args
        )

    def is_disconnect(self, e, connection, cursor):
        if super().is_disconnect(e, connection, cursor):
            return True
        elif isinstance(e, self.dbapi.Error):
            str_e = str(e).lower()
            return (
                "already closed" in str_e or "connection was killed" in str_e
            )
        else:
            return False

    def _extract_error_code(self, exception):
        if isinstance(exception.args[0], Exception):
            exception = exception.args[0]
        return exception.args[0]


dialect = MySQLDialect_pymysql

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\engine\_py_processors.py ===
# engine/_py_processors.py
# Copyright (C) 2010-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
# Copyright (C) 2010 Gaetan de Menten gdementen@gmail.com
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

"""defines generic type conversion functions, as used in bind and result
processors.

They all share one common characteristic: None is passed through unchanged.

"""

from __future__ import annotations

import datetime
from datetime import date as date_cls
from datetime import datetime as datetime_cls
from datetime import time as time_cls
from decimal import Decimal
import typing
from typing import Any
from typing import Callable
from typing import Optional
from typing import Type
from typing import TypeVar
from typing import Union


_DT = TypeVar(
    "_DT", bound=Union[datetime.datetime, datetime.time, datetime.date]
)


def str_to_datetime_processor_factory(
    regexp: typing.Pattern[str], type_: Callable[..., _DT]
) -> Callable[[Optional[str]], Optional[_DT]]:
    rmatch = regexp.match
    # Even on python2.6 datetime.strptime is both slower than this code
    # and it does not support microseconds.
    has_named_groups = bool(regexp.groupindex)

    def process(value: Optional[str]) -> Optional[_DT]:
        if value is None:
            return None
        else:
            try:
                m = rmatch(value)
            except TypeError as err:
                raise ValueError(
                    "Couldn't parse %s string '%r' "
                    "- value is not a string." % (type_.__name__, value)
                ) from err

            if m is None:
                raise ValueError(
                    "Couldn't parse %s string: "
                    "'%s'" % (type_.__name__, value)
                )
            if has_named_groups:
                groups = m.groupdict(0)
                return type_(
                    **dict(
                        list(
                            zip(
                                iter(groups.keys()),
                                list(map(int, iter(groups.values()))),
                            )
                        )
                    )
                )
            else:
                return type_(*list(map(int, m.groups(0))))

    return process


def to_decimal_processor_factory(
    target_class: Type[Decimal], scale: int
) -> Callable[[Optional[float]], Optional[Decimal]]:
    fstring = "%%.%df" % scale

    def process(value: Optional[float]) -> Optional[Decimal]:
        if value is None:
            return None
        else:
            return target_class(fstring % value)

    return process


def to_float(value: Optional[Union[int, float]]) -> Optional[float]:
    if value is None:
        return None
    else:
        return float(value)


def to_str(value: Optional[Any]) -> Optional[str]:
    if value is None:
        return None
    else:
        return str(value)


def int_to_boolean(value: Optional[int]) -> Optional[bool]:
    if value is None:
        return None
    else:
        return bool(value)


def str_to_datetime(value: Optional[str]) -> Optional[datetime.datetime]:
    if value is not None:
        dt_value = datetime_cls.fromisoformat(value)
    else:
        dt_value = None
    return dt_value


def str_to_time(value: Optional[str]) -> Optional[datetime.time]:
    if value is not None:
        dt_value = time_cls.fromisoformat(value)
    else:
        dt_value = None
    return dt_value


def str_to_date(value: Optional[str]) -> Optional[datetime.date]:
    if value is not None:
        dt_value = date_cls.fromisoformat(value)
    else:
        dt_value = None
    return dt_value

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\h11\_util.py ===
from typing import Any, Dict, NoReturn, Pattern, Tuple, Type, TypeVar, Union

__all__ = [
    "ProtocolError",
    "LocalProtocolError",
    "RemoteProtocolError",
    "validate",
    "bytesify",
]


class ProtocolError(Exception):
    """Exception indicating a violation of the HTTP/1.1 protocol.

    This as an abstract base class, with two concrete base classes:
    :exc:`LocalProtocolError`, which indicates that you tried to do something
    that HTTP/1.1 says is illegal, and :exc:`RemoteProtocolError`, which
    indicates that the remote peer tried to do something that HTTP/1.1 says is
    illegal. See :ref:`error-handling` for details.

    In addition to the normal :exc:`Exception` features, it has one attribute:

    .. attribute:: error_status_hint

       This gives a suggestion as to what status code a server might use if
       this error occurred as part of a request.

       For a :exc:`RemoteProtocolError`, this is useful as a suggestion for
       how you might want to respond to a misbehaving peer, if you're
       implementing a server.

       For a :exc:`LocalProtocolError`, this can be taken as a suggestion for
       how your peer might have responded to *you* if h11 had allowed you to
       continue.

       The default is 400 Bad Request, a generic catch-all for protocol
       violations.

    """

    def __init__(self, msg: str, error_status_hint: int = 400) -> None:
        if type(self) is ProtocolError:
            raise TypeError("tried to directly instantiate ProtocolError")
        Exception.__init__(self, msg)
        self.error_status_hint = error_status_hint


# Strategy: there are a number of public APIs where a LocalProtocolError can
# be raised (send(), all the different event constructors, ...), and only one
# public API where RemoteProtocolError can be raised
# (receive_data()). Therefore we always raise LocalProtocolError internally,
# and then receive_data will translate this into a RemoteProtocolError.
#
# Internally:
#   LocalProtocolError is the generic "ProtocolError".
# Externally:
#   LocalProtocolError is for local errors and RemoteProtocolError is for
#   remote errors.
class LocalProtocolError(ProtocolError):
    def _reraise_as_remote_protocol_error(self) -> NoReturn:
        # After catching a LocalProtocolError, use this method to re-raise it
        # as a RemoteProtocolError. This method must be called from inside an
        # except: block.
        #
        # An easy way to get an equivalent RemoteProtocolError is just to
        # modify 'self' in place.
        self.__class__ = RemoteProtocolError  # type: ignore
        # But the re-raising is somewhat non-trivial -- you might think that
        # now that we've modified the in-flight exception object, that just
        # doing 'raise' to re-raise it would be enough. But it turns out that
        # this doesn't work, because Python tracks the exception type
        # (exc_info[0]) separately from the exception object (exc_info[1]),
        # and we only modified the latter. So we really do need to re-raise
        # the new type explicitly.
        # On py3, the traceback is part of the exception object, so our
        # in-place modification preserved it and we can just re-raise:
        raise self


class RemoteProtocolError(ProtocolError):
    pass


def validate(
    regex: Pattern[bytes], data: bytes, msg: str = "malformed data", *format_args: Any
) -> Dict[str, bytes]:
    match = regex.fullmatch(data)
    if not match:
        if format_args:
            msg = msg.format(*format_args)
        raise LocalProtocolError(msg)
    return match.groupdict()


# Sentinel values
#
# - Inherit identity-based comparison and hashing from object
# - Have a nice repr
# - Have a *bonus property*: type(sentinel) is sentinel
#
# The bonus property is useful if you want to take the return value from
# next_event() and do some sort of dispatch based on type(event).

_T_Sentinel = TypeVar("_T_Sentinel", bound="Sentinel")


class Sentinel(type):
    def __new__(
        cls: Type[_T_Sentinel],
        name: str,
        bases: Tuple[type, ...],
        namespace: Dict[str, Any],
        **kwds: Any
    ) -> _T_Sentinel:
        assert bases == (Sentinel,)
        v = super().__new__(cls, name, bases, namespace, **kwds)
        v.__class__ = v  # type: ignore
        return v

    def __repr__(self) -> str:
        return self.__name__


# Used for methods, request targets, HTTP versions, header names, and header
# values. Accepts ascii-strings, or bytes/bytearray/memoryview/..., and always
# returns bytes.
def bytesify(s: Union[bytes, bytearray, memoryview, int, str]) -> bytes:
    # Fast-path:
    if type(s) is bytes:
        return s
    if isinstance(s, str):
        s = s.encode("ascii")
    if isinstance(s, int):
        raise TypeError("expected bytes-like object, not int")
    return bytes(s)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\quantization\fusions\fusion_layernorm.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation. All rights reserved.
# Licensed under the MIT License. See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------
from __future__ import annotations

import onnx

from ..onnx_model import ONNXModel
from .fusion import Fusion


class FusionLayerNormalization(Fusion):
    def __init__(self, model: ONNXModel):
        super().__init__(model, "LayerNormalization", "ReduceMean")

    def fuse(
        self,
        reduce_mean_node: onnx.NodeProto,
        input_name_to_nodes: dict[str, list[onnx.NodeProto]],
        output_name_to_node: dict[str, onnx.NodeProto],
    ):
        """
        Interface function that tries to fuse a node sequence containing a ReduceMean node into a single
        LayerNormalization node.

              +----------------------+
              |                      |
              |                      v
          [Root] --> ReduceMean -->  Sub  --> Pow --> ReduceMean --> Add --> Sqrt --> Div --> Mul --> Add
                     (axis=2 or -1)  |      (Y=2)   (axis=2 or -1)  (E-6 or E-12 or 0) ^
                                     |                                                 |
                                     +-------------------------------------------------+

         It also handles cases of duplicated sub nodes exported from older version of PyTorch:

              +----------------------+
              |                      v
              |           +-------> Sub-----------------------------------------------+
              |           |                                                           |
              |           |                                                           v
          [Root] --> ReduceMean -->  Sub  --> Pow --> ReduceMean --> Add --> Sqrt --> Div  --> Mul --> Add
              |                      ^
              |                      |
              +----------------------+
        """
        children = self.model.get_children(reduce_mean_node, input_name_to_nodes)
        if len(children) == 0 or len(children) > 2:
            return

        root_input = reduce_mean_node.input[0]

        if children[0].op_type != "Sub" or children[0].input[0] != root_input:
            return

        if len(children) == 2:
            if children[1].op_type != "Sub" or children[1].input[0] != root_input:
                return

        div_node = None
        for child in children:
            div_node = self.find_first_child_by_type(child, "Div", input_name_to_nodes, recursive=False)
            if div_node is not None:
                break
        if div_node is None:
            return

        path_id, parent_nodes, _ = self.match_parent_paths(
            div_node,
            [
                (["Sqrt", "Add", "ReduceMean", "Pow", "Sub"], [1, 0, 0, 0, 0]),
                (
                    ["Sqrt", "Add", "ReduceMean", "Pow", "Cast", "Sub"],
                    [1, 0, 0, 0, 0, 0],
                ),
            ],
            output_name_to_node,
        )
        if path_id < 0:
            return

        sub_node = parent_nodes[-1]
        if sub_node not in children:
            return

        second_add_node = parent_nodes[1]
        i, add_weight = self.get_constant_input(second_add_node)
        if add_weight is None or add_weight <= 0 or add_weight > 1.0e-4:
            # Skip fusion since epsilon value is not expected.
            return

        pow_node = parent_nodes[3]
        if self.find_constant_input(pow_node, 2.0) != 1:
            return

        mul_node = input_name_to_nodes[div_node.output[0]][0]
        if mul_node.op_type != "Mul":
            return

        last_add_node = input_name_to_nodes[mul_node.output[0]][0]
        if last_add_node.op_type != "Add":
            return

        subgraph_nodes = [reduce_mean_node]
        subgraph_nodes.extend(children)
        subgraph_nodes.extend(parent_nodes[:-1])

        subgraph_nodes.extend([last_add_node, mul_node, div_node])
        if not self.is_safe_to_fuse_nodes(
            subgraph_nodes,
            last_add_node.output,
            input_name_to_nodes,
            output_name_to_node,
        ):
            return

        weight_input = mul_node.input[1 - self.input_index(div_node.output[0], mul_node)]
        if not self.is_constant_with_specified_rank(weight_input, 1):
            return

        bias_input = last_add_node.input[1 - self.input_index(mul_node.output[0], last_add_node)]
        if not self.is_constant_with_specified_rank(bias_input, 1):
            return

        self.nodes_to_remove.extend(subgraph_nodes)

        normalize_node = onnx.helper.make_node(
            "LayerNormalization",
            name=self.create_unique_node_name(),
            inputs=[reduce_mean_node.input[0], weight_input, bias_input],
            outputs=[last_add_node.output[0]],
        )
        normalize_node.attribute.extend([onnx.helper.make_attribute("epsilon", float(add_weight))])
        self.nodes_to_add.append(normalize_node)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\testing\asyncio.py ===
# testing/asyncio.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors


# functions and wrappers to run tests, fixtures, provisioning and
# setup/teardown in an asyncio event loop, conditionally based on the
# current DB driver being used for a test.

# note that SQLAlchemy's asyncio integration also supports a method
# of running individual asyncio functions inside of separate event loops
# using "async_fallback" mode; however running whole functions in the event
# loop is a more accurate test for how SQLAlchemy's asyncio features
# would run in the real world.


from __future__ import annotations

from functools import wraps
import inspect

from . import config
from ..util.concurrency import _AsyncUtil

# may be set to False if the
# --disable-asyncio flag is passed to the test runner.
ENABLE_ASYNCIO = True
_async_util = _AsyncUtil()  # it has lazy init so just always create one


def _shutdown():
    """called when the test finishes"""
    _async_util.close()


def _run_coroutine_function(fn, *args, **kwargs):
    return _async_util.run(fn, *args, **kwargs)


def _assume_async(fn, *args, **kwargs):
    """Run a function in an asyncio loop unconditionally.

    This function is used for provisioning features like
    testing a database connection for server info.

    Note that for blocking IO database drivers, this means they block the
    event loop.

    """

    if not ENABLE_ASYNCIO:
        return fn(*args, **kwargs)

    return _async_util.run_in_greenlet(fn, *args, **kwargs)


def _maybe_async_provisioning(fn, *args, **kwargs):
    """Run a function in an asyncio loop if any current drivers might need it.

    This function is used for provisioning features that take
    place outside of a specific database driver being selected, so if the
    current driver that happens to be used for the provisioning operation
    is an async driver, it will run in asyncio and not fail.

    Note that for blocking IO database drivers, this means they block the
    event loop.

    """
    if not ENABLE_ASYNCIO:
        return fn(*args, **kwargs)

    if config.any_async:
        return _async_util.run_in_greenlet(fn, *args, **kwargs)
    else:
        return fn(*args, **kwargs)


def _maybe_async(fn, *args, **kwargs):
    """Run a function in an asyncio loop if the current selected driver is
    async.

    This function is used for test setup/teardown and tests themselves
    where the current DB driver is known.


    """
    if not ENABLE_ASYNCIO:
        return fn(*args, **kwargs)

    is_async = config._current.is_async

    if is_async:
        return _async_util.run_in_greenlet(fn, *args, **kwargs)
    else:
        return fn(*args, **kwargs)


def _maybe_async_wrapper(fn):
    """Apply the _maybe_async function to an existing function and return
    as a wrapped callable, supporting generator functions as well.

    This is currently used for pytest fixtures that support generator use.

    """

    if inspect.isgeneratorfunction(fn):
        _stop = object()

        def call_next(gen):
            try:
                return next(gen)
                # can't raise StopIteration in an awaitable.
            except StopIteration:
                return _stop

        @wraps(fn)
        def wrap_fixture(*args, **kwargs):
            gen = fn(*args, **kwargs)
            while True:
                value = _maybe_async(call_next, gen)
                if value is _stop:
                    break
                yield value

    else:

        @wraps(fn)
        def wrap_fixture(*args, **kwargs):
            return _maybe_async(fn, *args, **kwargs)

    return wrap_fixture

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\quantum\tests\test_spin.py ===
from sympy.concrete.summations import Sum
from sympy.core.function import expand
from sympy.core.numbers import (I, Rational, pi)
from sympy.core.singleton import S
from sympy.core.symbol import symbols
from sympy.functions.elementary.exponential import exp
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.trigonometric import (cos, sin)
from sympy.matrices.dense import Matrix
from sympy.abc import alpha, beta, gamma, j, m
from sympy.simplify import simplify

from sympy.physics.quantum import hbar, represent, Commutator, InnerProduct
from sympy.physics.quantum.qapply import qapply
from sympy.physics.quantum.tensorproduct import TensorProduct
from sympy.physics.quantum.cg import CG
from sympy.physics.quantum.spin import (
    Jx, Jy, Jz, Jplus, Jminus, J2,
    JxBra, JyBra, JzBra,
    JxKet, JyKet, JzKet,
    JxKetCoupled, JyKetCoupled, JzKetCoupled,
    couple, uncouple,
    Rotation, WignerD
)

from sympy.testing.pytest import raises, slow

j1, j2, j3, j4, m1, m2, m3, m4 = symbols('j1:5 m1:5')
j12, j13, j24, j34, j123, j134, mi, mi1, mp = symbols(
    'j12 j13 j24 j34 j123 j134 mi mi1 mp')


def assert_simplify_expand(e1, e2):
    """Helper for simplifying and expanding results.

    This is needed to help us test complex expressions whose form
    might change in subtle ways as the rest of sympy evolves.
    """
    assert simplify(e1.expand(tensorproduct=True)) == \
        simplify(e2.expand(tensorproduct=True))


def test_represent_spin_operators():
    assert represent(Jx) == hbar*Matrix([[0, 1], [1, 0]])/2
    assert represent(
        Jx, j=1) == hbar*sqrt(2)*Matrix([[0, 1, 0], [1, 0, 1], [0, 1, 0]])/2
    assert represent(Jy) == hbar*I*Matrix([[0, -1], [1, 0]])/2
    assert represent(Jy, j=1) == hbar*I*sqrt(2)*Matrix([[0, -1, 0], [1,
                     0, -1], [0, 1, 0]])/2
    assert represent(Jz) == hbar*Matrix([[1, 0], [0, -1]])/2
    assert represent(
        Jz, j=1) == hbar*Matrix([[1, 0, 0], [0, 0, 0], [0, 0, -1]])


def test_represent_spin_states():
    # Jx basis
    assert represent(JxKet(S.Half, S.Half), basis=Jx) == Matrix([1, 0])
    assert represent(JxKet(S.Half, Rational(-1, 2)), basis=Jx) == Matrix([0, 1])
    assert represent(JxKet(1, 1), basis=Jx) == Matrix([1, 0, 0])
    assert represent(JxKet(1, 0), basis=Jx) == Matrix([0, 1, 0])
    assert represent(JxKet(1, -1), basis=Jx) == Matrix([0, 0, 1])
    assert represent(
        JyKet(S.Half, S.Half), basis=Jx) == Matrix([exp(-I*pi/4), 0])
    assert represent(
        JyKet(S.Half, Rational(-1, 2)), basis=Jx) == Matrix([0, exp(I*pi/4)])
    assert represent(JyKet(1, 1), basis=Jx) == Matrix([-I, 0, 0])
    assert represent(JyKet(1, 0), basis=Jx) == Matrix([0, 1, 0])
    assert represent(JyKet(1, -1), basis=Jx) == Matrix([0, 0, I])
    assert represent(
        JzKet(S.Half, S.Half), basis=Jx) == sqrt(2)*Matrix([-1, 1])/2
    assert represent(
        JzKet(S.Half, Rational(-1, 2)), basis=Jx) == sqrt(2)*Matrix([-1, -1])/2
    assert represent(JzKet(1, 1), basis=Jx) == Matrix([1, -sqrt(2), 1])/2
    assert represent(JzKet(1, 0), basis=Jx) == sqrt(2)*Matrix([1, 0, -1])/2
    assert represent(JzKet(1, -1), basis=Jx) == Matrix([1, sqrt(2), 1])/2
    # Jy basis
    assert represent(
        JxKet(S.Half, S.Half), basis=Jy) == Matrix([exp(I*pi*Rational(-3, 4)), 0])
    assert represent(
        JxKet(S.Half, Rational(-1, 2)), basis=Jy) == Matrix([0, exp(I*pi*Rational(3, 4))])
    assert represent(JxKet(1, 1), basis=Jy) == Matrix([I, 0, 0])
    assert represent(JxKet(1, 0), basis=Jy) == Matrix([0, 1, 0])
    assert represent(JxKet(1, -1), basis=Jy) == Matrix([0, 0, -I])
    assert represent(JyKet(S.Half, S.Half), basis=Jy) == Matrix([1, 0])
    assert represent(JyKet(S.Half, Rational(-1, 2)), basis=Jy) == Matrix([0, 1])
    assert represent(JyKet(1, 1), basis=Jy) == Matrix([1, 0, 0])
    assert represent(JyKet(1, 0), basis=Jy) == Matrix([0, 1, 0])
    assert represent(JyKet(1, -1), basis=Jy) == Matrix([0, 0, 1])
    assert represent(
        JzKet(S.Half, S.Half), basis=Jy) == sqrt(2)*Matrix([-1, I])/2
    assert represent(
        JzKet(S.Half, Rational(-1, 2)), basis=Jy) == sqrt(2)*Matrix([I, -1])/2
    assert represent(JzKet(1, 1), basis=Jy) == Matrix([1, -I*sqrt(2), -1])/2
    assert represent(
        JzKet(1, 0), basis=Jy) == Matrix([-sqrt(2)*I, 0, -sqrt(2)*I])/2
    assert represent(JzKet(1, -1), basis=Jy) == Matrix([-1, -sqrt(2)*I, 1])/2
    # Jz basis
    assert represent(
        JxKet(S.Half, S.Half), basis=Jz) == sqrt(2)*Matrix([1, 1])/2
    assert represent(
        JxKet(S.Half, Rational(-1, 2)), basis=Jz) == sqrt(2)*Matrix([-1, 1])/2
    assert represent(JxKet(1, 1), basis=Jz) == Matrix([1, sqrt(2), 1])/2
    assert represent(JxKet(1, 0), basis=Jz) == sqrt(2)*Matrix([-1, 0, 1])/2
    assert represent(JxKet(1, -1), basis=Jz) == Matrix([1, -sqrt(2), 1])/2
    assert represent(
        JyKet(S.Half, S.Half), basis=Jz) == sqrt(2)*Matrix([-1, -I])/2
    assert represent(
        JyKet(S.Half, Rational(-1, 2)), basis=Jz) == sqrt(2)*Matrix([-I, -1])/2
    assert represent(JyKet(1, 1), basis=Jz) == Matrix([1, sqrt(2)*I, -1])/2
    assert represent(JyKet(1, 0), basis=Jz) == sqrt(2)*Matrix([I, 0, I])/2
    assert represent(JyKet(1, -1), basis=Jz) == Matrix([-1, sqrt(2)*I, 1])/2
    assert represent(JzKet(S.Half, S.Half), basis=Jz) == Matrix([1, 0])
    assert represent(JzKet(S.Half, Rational(-1, 2)), basis=Jz) == Matrix([0, 1])
    assert represent(JzKet(1, 1), basis=Jz) == Matrix([1, 0, 0])
    assert represent(JzKet(1, 0), basis=Jz) == Matrix([0, 1, 0])
    assert represent(JzKet(1, -1), basis=Jz) == Matrix([0, 0, 1])


def test_represent_uncoupled_states():
    # Jx basis
    assert represent(TensorProduct(JxKet(S.Half, S.Half), JxKet(S.Half, S.Half)), basis=Jx) == \
        Matrix([1, 0, 0, 0])
    assert represent(TensorProduct(JxKet(S.Half, S.Half), JxKet(S.Half, Rational(-1, 2))), basis=Jx) == \
        Matrix([0, 1, 0, 0])
    assert represent(TensorProduct(JxKet(S.Half, Rational(-1, 2)), JxKet(S.Half, S.Half)), basis=Jx) == \
        Matrix([0, 0, 1, 0])
    assert represent(TensorProduct(JxKet(S.Half, Rational(-1, 2)), JxKet(S.Half, Rational(-1, 2))), basis=Jx) == \
        Matrix([0, 0, 0, 1])
    assert represent(TensorProduct(JyKet(S.Half, S.Half), JyKet(S.Half, S.Half)), basis=Jx) == \
        Matrix([-I, 0, 0, 0])
    assert represent(TensorProduct(JyKet(S.Half, S.Half), JyKet(S.Half, Rational(-1, 2))), basis=Jx) == \
        Matrix([0, 1, 0, 0])
    assert represent(TensorProduct(JyKet(S.Half, Rational(-1, 2)), JyKet(S.Half, S.Half)), basis=Jx) == \
        Matrix([0, 0, 1, 0])
    assert represent(TensorProduct(JyKet(S.Half, Rational(-1, 2)), JyKet(S.Half, Rational(-1, 2))), basis=Jx) == \
        Matrix([0, 0, 0, I])
    assert represent(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), basis=Jx) == \
        Matrix([S.Half, Rational(-1, 2), Rational(-1, 2), S.Half])
    assert represent(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), basis=Jx) == \
        Matrix([S.Half, S.Half, Rational(-1, 2), Rational(-1, 2)])
    assert represent(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), basis=Jx) == \
        Matrix([S.Half, Rational(-1, 2), S.Half, Rational(-1, 2)])
    assert represent(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), basis=Jx) == \
        Matrix([S.Half, S.Half, S.Half, S.Half])
    # Jy basis
    assert represent(TensorProduct(JxKet(S.Half, S.Half), JxKet(S.Half, S.Half)), basis=Jy) == \
        Matrix([I, 0, 0, 0])
    assert represent(TensorProduct(JxKet(S.Half, S.Half), JxKet(S.Half, Rational(-1, 2))), basis=Jy) == \
        Matrix([0, 1, 0, 0])
    assert represent(TensorProduct(JxKet(S.Half, Rational(-1, 2)), JxKet(S.Half, S.Half)), basis=Jy) == \
        Matrix([0, 0, 1, 0])
    assert represent(TensorProduct(JxKet(S.Half, Rational(-1, 2)), JxKet(S.Half, Rational(-1, 2))), basis=Jy) == \
        Matrix([0, 0, 0, -I])
    assert represent(TensorProduct(JyKet(S.Half, S.Half), JyKet(S.Half, S.Half)), basis=Jy) == \
        Matrix([1, 0, 0, 0])
    assert represent(TensorProduct(JyKet(S.Half, S.Half), JyKet(S.Half, Rational(-1, 2))), basis=Jy) == \
        Matrix([0, 1, 0, 0])
    assert represent(TensorProduct(JyKet(S.Half, Rational(-1, 2)), JyKet(S.Half, S.Half)), basis=Jy) == \
        Matrix([0, 0, 1, 0])
    assert represent(TensorProduct(JyKet(S.Half, Rational(-1, 2)), JyKet(S.Half, Rational(-1, 2))), basis=Jy) == \
        Matrix([0, 0, 0, 1])
    assert represent(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), basis=Jy) == \
        Matrix([S.Half, -I/2, -I/2, Rational(-1, 2)])
    assert represent(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), basis=Jy) == \
        Matrix([-I/2, S.Half, Rational(-1, 2), -I/2])
    assert represent(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), basis=Jy) == \
        Matrix([-I/2, Rational(-1, 2), S.Half, -I/2])
    assert represent(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), basis=Jy) == \
        Matrix([Rational(-1, 2), -I/2, -I/2, S.Half])
    # Jz basis
    assert represent(TensorProduct(JxKet(S.Half, S.Half), JxKet(S.Half, S.Half)), basis=Jz) == \
        Matrix([S.Half, S.Half, S.Half, S.Half])
    assert represent(TensorProduct(JxKet(S.Half, S.Half), JxKet(S.Half, Rational(-1, 2))), basis=Jz) == \
        Matrix([Rational(-1, 2), S.Half, Rational(-1, 2), S.Half])
    assert represent(TensorProduct(JxKet(S.Half, Rational(-1, 2)), JxKet(S.Half, S.Half)), basis=Jz) == \
        Matrix([Rational(-1, 2), Rational(-1, 2), S.Half, S.Half])
    assert represent(TensorProduct(JxKet(S.Half, Rational(-1, 2)), JxKet(S.Half, Rational(-1, 2))), basis=Jz) == \
        Matrix([S.Half, Rational(-1, 2), Rational(-1, 2), S.Half])
    assert represent(TensorProduct(JyKet(S.Half, S.Half), JyKet(S.Half, S.Half)), basis=Jz) == \
        Matrix([S.Half, I/2, I/2, Rational(-1, 2)])
    assert represent(TensorProduct(JyKet(S.Half, S.Half), JyKet(S.Half, Rational(-1, 2))), basis=Jz) == \
        Matrix([I/2, S.Half, Rational(-1, 2), I/2])
    assert represent(TensorProduct(JyKet(S.Half, Rational(-1, 2)), JyKet(S.Half, S.Half)), basis=Jz) == \
        Matrix([I/2, Rational(-1, 2), S.Half, I/2])
    assert represent(TensorProduct(JyKet(S.Half, Rational(-1, 2)), JyKet(S.Half, Rational(-1, 2))), basis=Jz) == \
        Matrix([Rational(-1, 2), I/2, I/2, S.Half])
    assert represent(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), basis=Jz) == \
        Matrix([1, 0, 0, 0])
    assert represent(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), basis=Jz) == \
        Matrix([0, 1, 0, 0])
    assert represent(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), basis=Jz) == \
        Matrix([0, 0, 1, 0])
    assert represent(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), basis=Jz) == \
        Matrix([0, 0, 0, 1])


def test_represent_coupled_states():
    # Jx basis
    assert represent(JxKetCoupled(0, 0, (S.Half, S.Half)), basis=Jx) == \
        Matrix([1, 0, 0, 0])
    assert represent(JxKetCoupled(1, 1, (S.Half, S.Half)), basis=Jx) == \
        Matrix([0, 1, 0, 0])
    assert represent(JxKetCoupled(1, 0, (S.Half, S.Half)), basis=Jx) == \
        Matrix([0, 0, 1, 0])
    assert represent(JxKetCoupled(1, -1, (S.Half, S.Half)), basis=Jx) == \
        Matrix([0, 0, 0, 1])
    assert represent(JyKetCoupled(0, 0, (S.Half, S.Half)), basis=Jx) == \
        Matrix([1, 0, 0, 0])
    assert represent(JyKetCoupled(1, 1, (S.Half, S.Half)), basis=Jx) == \
        Matrix([0, -I, 0, 0])
    assert represent(JyKetCoupled(1, 0, (S.Half, S.Half)), basis=Jx) == \
        Matrix([0, 0, 1, 0])
    assert represent(JyKetCoupled(1, -1, (S.Half, S.Half)), basis=Jx) == \
        Matrix([0, 0, 0, I])
    assert represent(JzKetCoupled(0, 0, (S.Half, S.Half)), basis=Jx) == \
        Matrix([1, 0, 0, 0])
    assert represent(JzKetCoupled(1, 1, (S.Half, S.Half)), basis=Jx) == \
        Matrix([0, S.Half, -sqrt(2)/2, S.Half])
    assert represent(JzKetCoupled(1, 0, (S.Half, S.Half)), basis=Jx) == \
        Matrix([0, sqrt(2)/2, 0, -sqrt(2)/2])
    assert represent(JzKetCoupled(1, -1, (S.Half, S.Half)), basis=Jx) == \
        Matrix([0, S.Half, sqrt(2)/2, S.Half])
    # Jy basis
    assert represent(JxKetCoupled(0, 0, (S.Half, S.Half)), basis=Jy) == \
        Matrix([1, 0, 0, 0])
    assert represent(JxKetCoupled(1, 1, (S.Half, S.Half)), basis=Jy) == \
        Matrix([0, I, 0, 0])
    assert represent(JxKetCoupled(1, 0, (S.Half, S.Half)), basis=Jy) == \
        Matrix([0, 0, 1, 0])
    assert represent(JxKetCoupled(1, -1, (S.Half, S.Half)), basis=Jy) == \
        Matrix([0, 0, 0, -I])
    assert represent(JyKetCoupled(0, 0, (S.Half, S.Half)), basis=Jy) == \
        Matrix([1, 0, 0, 0])
    assert represent(JyKetCoupled(1, 1, (S.Half, S.Half)), basis=Jy) == \
        Matrix([0, 1, 0, 0])
    assert represent(JyKetCoupled(1, 0, (S.Half, S.Half)), basis=Jy) == \
        Matrix([0, 0, 1, 0])
    assert represent(JyKetCoupled(1, -1, (S.Half, S.Half)), basis=Jy) == \
        Matrix([0, 0, 0, 1])
    assert represent(JzKetCoupled(0, 0, (S.Half, S.Half)), basis=Jy) == \
        Matrix([1, 0, 0, 0])
    assert represent(JzKetCoupled(1, 1, (S.Half, S.Half)), basis=Jy) == \
        Matrix([0, S.Half, -I*sqrt(2)/2, Rational(-1, 2)])
    assert represent(JzKetCoupled(1, 0, (S.Half, S.Half)), basis=Jy) == \
        Matrix([0, -I*sqrt(2)/2, 0, -I*sqrt(2)/2])
    assert represent(JzKetCoupled(1, -1, (S.Half, S.Half)), basis=Jy) == \
        Matrix([0, Rational(-1, 2), -I*sqrt(2)/2, S.Half])
    # Jz basis
    assert represent(JxKetCoupled(0, 0, (S.Half, S.Half)), basis=Jz) == \
        Matrix([1, 0, 0, 0])
    assert represent(JxKetCoupled(1, 1, (S.Half, S.Half)), basis=Jz) == \
        Matrix([0, S.Half, sqrt(2)/2, S.Half])
    assert represent(JxKetCoupled(1, 0, (S.Half, S.Half)), basis=Jz) == \
        Matrix([0, -sqrt(2)/2, 0, sqrt(2)/2])
    assert represent(JxKetCoupled(1, -1, (S.Half, S.Half)), basis=Jz) == \
        Matrix([0, S.Half, -sqrt(2)/2, S.Half])
    assert represent(JyKetCoupled(0, 0, (S.Half, S.Half)), basis=Jz) == \
        Matrix([1, 0, 0, 0])
    assert represent(JyKetCoupled(1, 1, (S.Half, S.Half)), basis=Jz) == \
        Matrix([0, S.Half, I*sqrt(2)/2, Rational(-1, 2)])
    assert represent(JyKetCoupled(1, 0, (S.Half, S.Half)), basis=Jz) == \
        Matrix([0, I*sqrt(2)/2, 0, I*sqrt(2)/2])
    assert represent(JyKetCoupled(1, -1, (S.Half, S.Half)), basis=Jz) == \
        Matrix([0, Rational(-1, 2), I*sqrt(2)/2, S.Half])
    assert represent(JzKetCoupled(0, 0, (S.Half, S.Half)), basis=Jz) == \
        Matrix([1, 0, 0, 0])
    assert represent(JzKetCoupled(1, 1, (S.Half, S.Half)), basis=Jz) == \
        Matrix([0, 1, 0, 0])
    assert represent(JzKetCoupled(1, 0, (S.Half, S.Half)), basis=Jz) == \
        Matrix([0, 0, 1, 0])
    assert represent(JzKetCoupled(1, -1, (S.Half, S.Half)), basis=Jz) == \
        Matrix([0, 0, 0, 1])


def test_represent_rotation():
    assert represent(Rotation(0, pi/2, 0)) == \
        Matrix(
            [[WignerD(
                S(
                    1)/2, S(
                        1)/2, S(
                            1)/2, 0, pi/2, 0), WignerD(
                                S.Half, S.Half, Rational(-1, 2), 0, pi/2, 0)],
                [WignerD(S.Half, Rational(-1, 2), S.Half, 0, pi/2, 0), WignerD(S.Half, Rational(-1, 2), Rational(-1, 2), 0, pi/2, 0)]])
    assert represent(Rotation(0, pi/2, 0), doit=True) == \
        Matrix([[sqrt(2)/2, -sqrt(2)/2],
                [sqrt(2)/2, sqrt(2)/2]])


def test_rewrite_same():
    # Rewrite to same basis
    assert JxBra(1, 1).rewrite('Jx') == JxBra(1, 1)
    assert JxBra(j, m).rewrite('Jx') == JxBra(j, m)
    assert JxKet(1, 1).rewrite('Jx') == JxKet(1, 1)
    assert JxKet(j, m).rewrite('Jx') == JxKet(j, m)


def test_rewrite_Bra():
    # Numerical
    assert JxBra(1, 1).rewrite('Jy') == -I*JyBra(1, 1)
    assert JxBra(1, 0).rewrite('Jy') == JyBra(1, 0)
    assert JxBra(1, -1).rewrite('Jy') == I*JyBra(1, -1)
    assert JxBra(1, 1).rewrite(
        'Jz') == JzBra(1, 1)/2 + JzBra(1, 0)/sqrt(2) + JzBra(1, -1)/2
    assert JxBra(
        1, 0).rewrite('Jz') == -sqrt(2)*JzBra(1, 1)/2 + sqrt(2)*JzBra(1, -1)/2
    assert JxBra(1, -1).rewrite(
        'Jz') == JzBra(1, 1)/2 - JzBra(1, 0)/sqrt(2) + JzBra(1, -1)/2
    assert JyBra(1, 1).rewrite('Jx') == I*JxBra(1, 1)
    assert JyBra(1, 0).rewrite('Jx') == JxBra(1, 0)
    assert JyBra(1, -1).rewrite('Jx') == -I*JxBra(1, -1)
    assert JyBra(1, 1).rewrite(
        'Jz') == JzBra(1, 1)/2 - sqrt(2)*I*JzBra(1, 0)/2 - JzBra(1, -1)/2
    assert JyBra(1, 0).rewrite(
        'Jz') == -sqrt(2)*I*JzBra(1, 1)/2 - sqrt(2)*I*JzBra(1, -1)/2
    assert JyBra(1, -1).rewrite(
        'Jz') == -JzBra(1, 1)/2 - sqrt(2)*I*JzBra(1, 0)/2 + JzBra(1, -1)/2
    assert JzBra(1, 1).rewrite(
        'Jx') == JxBra(1, 1)/2 - sqrt(2)*JxBra(1, 0)/2 + JxBra(1, -1)/2
    assert JzBra(
        1, 0).rewrite('Jx') == sqrt(2)*JxBra(1, 1)/2 - sqrt(2)*JxBra(1, -1)/2
    assert JzBra(1, -1).rewrite(
        'Jx') == JxBra(1, 1)/2 + sqrt(2)*JxBra(1, 0)/2 + JxBra(1, -1)/2
    assert JzBra(1, 1).rewrite(
        'Jy') == JyBra(1, 1)/2 + sqrt(2)*I*JyBra(1, 0)/2 - JyBra(1, -1)/2
    assert JzBra(1, 0).rewrite(
        'Jy') == sqrt(2)*I*JyBra(1, 1)/2 + sqrt(2)*I*JyBra(1, -1)/2
    assert JzBra(1, -1).rewrite(
        'Jy') == -JyBra(1, 1)/2 + sqrt(2)*I*JyBra(1, 0)/2 + JyBra(1, -1)/2
    # Symbolic
    assert JxBra(j, m).rewrite('Jy') == Sum(
        WignerD(j, mi, m, pi*Rational(3, 2), 0, 0) * JyBra(j, mi), (mi, -j, j))
    assert JxBra(j, m).rewrite('Jz') == Sum(
        WignerD(j, mi, m, 0, pi/2, 0) * JzBra(j, mi), (mi, -j, j))
    assert JyBra(j, m).rewrite('Jx') == Sum(
        WignerD(j, mi, m, 0, 0, pi/2) * JxBra(j, mi), (mi, -j, j))
    assert JyBra(j, m).rewrite('Jz') == Sum(
        WignerD(j, mi, m, pi*Rational(3, 2), -pi/2, pi/2) * JzBra(j, mi), (mi, -j, j))
    assert JzBra(j, m).rewrite('Jx') == Sum(
        WignerD(j, mi, m, 0, pi*Rational(3, 2), 0) * JxBra(j, mi), (mi, -j, j))
    assert JzBra(j, m).rewrite('Jy') == Sum(
        WignerD(j, mi, m, pi*Rational(3, 2), pi/2, pi/2) * JyBra(j, mi), (mi, -j, j))


def test_rewrite_Ket():
    # Numerical
    assert JxKet(1, 1).rewrite('Jy') == I*JyKet(1, 1)
    assert JxKet(1, 0).rewrite('Jy') == JyKet(1, 0)
    assert JxKet(1, -1).rewrite('Jy') == -I*JyKet(1, -1)
    assert JxKet(1, 1).rewrite(
        'Jz') == JzKet(1, 1)/2 + JzKet(1, 0)/sqrt(2) + JzKet(1, -1)/2
    assert JxKet(
        1, 0).rewrite('Jz') == -sqrt(2)*JzKet(1, 1)/2 + sqrt(2)*JzKet(1, -1)/2
    assert JxKet(1, -1).rewrite(
        'Jz') == JzKet(1, 1)/2 - JzKet(1, 0)/sqrt(2) + JzKet(1, -1)/2
    assert JyKet(1, 1).rewrite('Jx') == -I*JxKet(1, 1)
    assert JyKet(1, 0).rewrite('Jx') == JxKet(1, 0)
    assert JyKet(1, -1).rewrite('Jx') == I*JxKet(1, -1)
    assert JyKet(1, 1).rewrite(
        'Jz') == JzKet(1, 1)/2 + sqrt(2)*I*JzKet(1, 0)/2 - JzKet(1, -1)/2
    assert JyKet(1, 0).rewrite(
        'Jz') == sqrt(2)*I*JzKet(1, 1)/2 + sqrt(2)*I*JzKet(1, -1)/2
    assert JyKet(1, -1).rewrite(
        'Jz') == -JzKet(1, 1)/2 + sqrt(2)*I*JzKet(1, 0)/2 + JzKet(1, -1)/2
    assert JzKet(1, 1).rewrite(
        'Jx') == JxKet(1, 1)/2 - sqrt(2)*JxKet(1, 0)/2 + JxKet(1, -1)/2
    assert JzKet(
        1, 0).rewrite('Jx') == sqrt(2)*JxKet(1, 1)/2 - sqrt(2)*JxKet(1, -1)/2
    assert JzKet(1, -1).rewrite(
        'Jx') == JxKet(1, 1)/2 + sqrt(2)*JxKet(1, 0)/2 + JxKet(1, -1)/2
    assert JzKet(1, 1).rewrite(
        'Jy') == JyKet(1, 1)/2 - sqrt(2)*I*JyKet(1, 0)/2 - JyKet(1, -1)/2
    assert JzKet(1, 0).rewrite(
        'Jy') == -sqrt(2)*I*JyKet(1, 1)/2 - sqrt(2)*I*JyKet(1, -1)/2
    assert JzKet(1, -1).rewrite(
        'Jy') == -JyKet(1, 1)/2 - sqrt(2)*I*JyKet(1, 0)/2 + JyKet(1, -1)/2
    # Symbolic
    assert JxKet(j, m).rewrite('Jy') == Sum(
        WignerD(j, mi, m, pi*Rational(3, 2), 0, 0) * JyKet(j, mi), (mi, -j, j))
    assert JxKet(j, m).rewrite('Jz') == Sum(
        WignerD(j, mi, m, 0, pi/2, 0) * JzKet(j, mi), (mi, -j, j))
    assert JyKet(j, m).rewrite('Jx') == Sum(
        WignerD(j, mi, m, 0, 0, pi/2) * JxKet(j, mi), (mi, -j, j))
    assert JyKet(j, m).rewrite('Jz') == Sum(
        WignerD(j, mi, m, pi*Rational(3, 2), -pi/2, pi/2) * JzKet(j, mi), (mi, -j, j))
    assert JzKet(j, m).rewrite('Jx') == Sum(
        WignerD(j, mi, m, 0, pi*Rational(3, 2), 0) * JxKet(j, mi), (mi, -j, j))
    assert JzKet(j, m).rewrite('Jy') == Sum(
        WignerD(j, mi, m, pi*Rational(3, 2), pi/2, pi/2) * JyKet(j, mi), (mi, -j, j))


def test_rewrite_uncoupled_state():
    # Numerical
    assert TensorProduct(JyKet(1, 1), JxKet(
        1, 1)).rewrite('Jx') == -I*TensorProduct(JxKet(1, 1), JxKet(1, 1))
    assert TensorProduct(JyKet(1, 0), JxKet(
        1, 1)).rewrite('Jx') == TensorProduct(JxKet(1, 0), JxKet(1, 1))
    assert TensorProduct(JyKet(1, -1), JxKet(
        1, 1)).rewrite('Jx') == I*TensorProduct(JxKet(1, -1), JxKet(1, 1))
    assert TensorProduct(JzKet(1, 1), JxKet(1, 1)).rewrite('Jx') == \
        TensorProduct(JxKet(1, -1), JxKet(1, 1))/2 - sqrt(2)*TensorProduct(JxKet(
            1, 0), JxKet(1, 1))/2 + TensorProduct(JxKet(1, 1), JxKet(1, 1))/2
    assert TensorProduct(JzKet(1, 0), JxKet(1, 1)).rewrite('Jx') == \
        -sqrt(2)*TensorProduct(JxKet(1, -1), JxKet(1, 1))/2 + sqrt(
            2)*TensorProduct(JxKet(1, 1), JxKet(1, 1))/2
    assert TensorProduct(JzKet(1, -1), JxKet(1, 1)).rewrite('Jx') == \
        TensorProduct(JxKet(1, -1), JxKet(1, 1))/2 + sqrt(2)*TensorProduct(JxKet(1, 0), JxKet(1, 1))/2 + TensorProduct(JxKet(1, 1), JxKet(1, 1))/2
    assert TensorProduct(JxKet(1, 1), JyKet(
        1, 1)).rewrite('Jy') == I*TensorProduct(JyKet(1, 1), JyKet(1, 1))
    assert TensorProduct(JxKet(1, 0), JyKet(
        1, 1)).rewrite('Jy') == TensorProduct(JyKet(1, 0), JyKet(1, 1))
    assert TensorProduct(JxKet(1, -1), JyKet(
        1, 1)).rewrite('Jy') == -I*TensorProduct(JyKet(1, -1), JyKet(1, 1))
    assert TensorProduct(JzKet(1, 1), JyKet(1, 1)).rewrite('Jy') == \
        -TensorProduct(JyKet(1, -1), JyKet(1, 1))/2 - sqrt(2)*I*TensorProduct(JyKet(1, 0), JyKet(1, 1))/2 + TensorProduct(JyKet(1, 1), JyKet(1, 1))/2
    assert TensorProduct(JzKet(1, 0), JyKet(1, 1)).rewrite('Jy') == \
        -sqrt(2)*I*TensorProduct(JyKet(1, -1), JyKet(
            1, 1))/2 - sqrt(2)*I*TensorProduct(JyKet(1, 1), JyKet(1, 1))/2
    assert TensorProduct(JzKet(1, -1), JyKet(1, 1)).rewrite('Jy') == \
        TensorProduct(JyKet(1, -1), JyKet(1, 1))/2 - sqrt(2)*I*TensorProduct(JyKet(1, 0), JyKet(1, 1))/2 - TensorProduct(JyKet(1, 1), JyKet(1, 1))/2
    assert TensorProduct(JxKet(1, 1), JzKet(1, 1)).rewrite('Jz') == \
        TensorProduct(JzKet(1, -1), JzKet(1, 1))/2 + sqrt(2)*TensorProduct(JzKet(1, 0), JzKet(1, 1))/2 + TensorProduct(JzKet(1, 1), JzKet(1, 1))/2
    assert TensorProduct(JxKet(1, 0), JzKet(1, 1)).rewrite('Jz') == \
        sqrt(2)*TensorProduct(JzKet(1, -1), JzKet(
            1, 1))/2 - sqrt(2)*TensorProduct(JzKet(1, 1), JzKet(1, 1))/2
    assert TensorProduct(JxKet(1, -1), JzKet(1, 1)).rewrite('Jz') == \
        TensorProduct(JzKet(1, -1), JzKet(1, 1))/2 - sqrt(2)*TensorProduct(JzKet(1, 0), JzKet(1, 1))/2 + TensorProduct(JzKet(1, 1), JzKet(1, 1))/2
    assert TensorProduct(JyKet(1, 1), JzKet(1, 1)).rewrite('Jz') == \
        -TensorProduct(JzKet(1, -1), JzKet(1, 1))/2 + sqrt(2)*I*TensorProduct(JzKet(1, 0), JzKet(1, 1))/2 + TensorProduct(JzKet(1, 1), JzKet(1, 1))/2
    assert TensorProduct(JyKet(1, 0), JzKet(1, 1)).rewrite('Jz') == \
        sqrt(2)*I*TensorProduct(JzKet(1, -1), JzKet(
            1, 1))/2 + sqrt(2)*I*TensorProduct(JzKet(1, 1), JzKet(1, 1))/2
    assert TensorProduct(JyKet(1, -1), JzKet(1, 1)).rewrite('Jz') == \
        TensorProduct(JzKet(1, -1), JzKet(1, 1))/2 + sqrt(2)*I*TensorProduct(JzKet(1, 0), JzKet(1, 1))/2 - TensorProduct(JzKet(1, 1), JzKet(1, 1))/2
    # Symbolic
    assert TensorProduct(JyKet(j1, m1), JxKet(j2, m2)).rewrite('Jy') == \
        TensorProduct(JyKet(j1, m1), Sum(
            WignerD(j2, mi, m2, pi*Rational(3, 2), 0, 0) * JyKet(j2, mi), (mi, -j2, j2)))
    assert TensorProduct(JzKet(j1, m1), JxKet(j2, m2)).rewrite('Jz') == \
        TensorProduct(JzKet(j1, m1), Sum(
            WignerD(j2, mi, m2, 0, pi/2, 0) * JzKet(j2, mi), (mi, -j2, j2)))
    assert TensorProduct(JxKet(j1, m1), JyKet(j2, m2)).rewrite('Jx') == \
        TensorProduct(JxKet(j1, m1), Sum(
            WignerD(j2, mi, m2, 0, 0, pi/2) * JxKet(j2, mi), (mi, -j2, j2)))
    assert TensorProduct(JzKet(j1, m1), JyKet(j2, m2)).rewrite('Jz') == \
        TensorProduct(JzKet(j1, m1), Sum(WignerD(
            j2, mi, m2, pi*Rational(3, 2), -pi/2, pi/2) * JzKet(j2, mi), (mi, -j2, j2)))
    assert TensorProduct(JxKet(j1, m1), JzKet(j2, m2)).rewrite('Jx') == \
        TensorProduct(JxKet(j1, m1), Sum(
            WignerD(j2, mi, m2, 0, pi*Rational(3, 2), 0) * JxKet(j2, mi), (mi, -j2, j2)))
    assert TensorProduct(JyKet(j1, m1), JzKet(j2, m2)).rewrite('Jy') == \
        TensorProduct(JyKet(j1, m1), Sum(WignerD(
            j2, mi, m2, pi*Rational(3, 2), pi/2, pi/2) * JyKet(j2, mi), (mi, -j2, j2)))


def test_rewrite_coupled_state():
    # Numerical
    assert JyKetCoupled(0, 0, (S.Half, S.Half)).rewrite('Jx') == \
        JxKetCoupled(0, 0, (S.Half, S.Half))
    assert JyKetCoupled(1, 1, (S.Half, S.Half)).rewrite('Jx') == \
        -I*JxKetCoupled(1, 1, (S.Half, S.Half))
    assert JyKetCoupled(1, 0, (S.Half, S.Half)).rewrite('Jx') == \
        JxKetCoupled(1, 0, (S.Half, S.Half))
    assert JyKetCoupled(1, -1, (S.Half, S.Half)).rewrite('Jx') == \
        I*JxKetCoupled(1, -1, (S.Half, S.Half))
    assert JzKetCoupled(0, 0, (S.Half, S.Half)).rewrite('Jx') == \
        JxKetCoupled(0, 0, (S.Half, S.Half))
    assert JzKetCoupled(1, 1, (S.Half, S.Half)).rewrite('Jx') == \
        JxKetCoupled(1, 1, (S.Half, S.Half))/2 - sqrt(2)*JxKetCoupled(1, 0, (
            S.Half, S.Half))/2 + JxKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JzKetCoupled(1, 0, (S.Half, S.Half)).rewrite('Jx') == \
        sqrt(2)*JxKetCoupled(1, 1, (S(
            1)/2, S.Half))/2 - sqrt(2)*JxKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JzKetCoupled(1, -1, (S.Half, S.Half)).rewrite('Jx') == \
        JxKetCoupled(1, 1, (S.Half, S.Half))/2 + sqrt(2)*JxKetCoupled(1, 0, (
            S.Half, S.Half))/2 + JxKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JxKetCoupled(0, 0, (S.Half, S.Half)).rewrite('Jy') == \
        JyKetCoupled(0, 0, (S.Half, S.Half))
    assert JxKetCoupled(1, 1, (S.Half, S.Half)).rewrite('Jy') == \
        I*JyKetCoupled(1, 1, (S.Half, S.Half))
    assert JxKetCoupled(1, 0, (S.Half, S.Half)).rewrite('Jy') == \
        JyKetCoupled(1, 0, (S.Half, S.Half))
    assert JxKetCoupled(1, -1, (S.Half, S.Half)).rewrite('Jy') == \
        -I*JyKetCoupled(1, -1, (S.Half, S.Half))
    assert JzKetCoupled(0, 0, (S.Half, S.Half)).rewrite('Jy') == \
        JyKetCoupled(0, 0, (S.Half, S.Half))
    assert JzKetCoupled(1, 1, (S.Half, S.Half)).rewrite('Jy') == \
        JyKetCoupled(1, 1, (S.Half, S.Half))/2 - I*sqrt(2)*JyKetCoupled(1, 0, (
            S.Half, S.Half))/2 - JyKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JzKetCoupled(1, 0, (S.Half, S.Half)).rewrite('Jy') == \
        -I*sqrt(2)*JyKetCoupled(1, 1, (S.Half, S.Half))/2 - I*sqrt(
            2)*JyKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JzKetCoupled(1, -1, (S.Half, S.Half)).rewrite('Jy') == \
        -JyKetCoupled(1, 1, (S.Half, S.Half))/2 - I*sqrt(2)*JyKetCoupled(1, 0, (S.Half, S.Half))/2 + JyKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JxKetCoupled(0, 0, (S.Half, S.Half)).rewrite('Jz') == \
        JzKetCoupled(0, 0, (S.Half, S.Half))
    assert JxKetCoupled(1, 1, (S.Half, S.Half)).rewrite('Jz') == \
        JzKetCoupled(1, 1, (S.Half, S.Half))/2 + sqrt(2)*JzKetCoupled(1, 0, (
            S.Half, S.Half))/2 + JzKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JxKetCoupled(1, 0, (S.Half, S.Half)).rewrite('Jz') == \
        -sqrt(2)*JzKetCoupled(1, 1, (S(
            1)/2, S.Half))/2 + sqrt(2)*JzKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JxKetCoupled(1, -1, (S.Half, S.Half)).rewrite('Jz') == \
        JzKetCoupled(1, 1, (S.Half, S.Half))/2 - sqrt(2)*JzKetCoupled(1, 0, (
            S.Half, S.Half))/2 + JzKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JyKetCoupled(0, 0, (S.Half, S.Half)).rewrite('Jz') == \
        JzKetCoupled(0, 0, (S.Half, S.Half))
    assert JyKetCoupled(1, 1, (S.Half, S.Half)).rewrite('Jz') == \
        JzKetCoupled(1, 1, (S.Half, S.Half))/2 + I*sqrt(2)*JzKetCoupled(1, 0, (
            S.Half, S.Half))/2 - JzKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JyKetCoupled(1, 0, (S.Half, S.Half)).rewrite('Jz') == \
        I*sqrt(2)*JzKetCoupled(1, 1, (S.Half, S.Half))/2 + I*sqrt(
            2)*JzKetCoupled(1, -1, (S.Half, S.Half))/2
    assert JyKetCoupled(1, -1, (S.Half, S.Half)).rewrite('Jz') == \
        -JzKetCoupled(1, 1, (S.Half, S.Half))/2 + I*sqrt(2)*JzKetCoupled(1, 0, (S.Half, S.Half))/2 + JzKetCoupled(1, -1, (S.Half, S.Half))/2
    # Symbolic
    assert JyKetCoupled(j, m, (j1, j2)).rewrite('Jx') == \
        Sum(WignerD(j, mi, m, 0, 0, pi/2) * JxKetCoupled(j, mi, (
            j1, j2)), (mi, -j, j))
    assert JzKetCoupled(j, m, (j1, j2)).rewrite('Jx') == \
        Sum(WignerD(j, mi, m, 0, pi*Rational(3, 2), 0) * JxKetCoupled(j, mi, (
            j1, j2)), (mi, -j, j))
    assert JxKetCoupled(j, m, (j1, j2)).rewrite('Jy') == \
        Sum(WignerD(j, mi, m, pi*Rational(3, 2), 0, 0) * JyKetCoupled(j, mi, (
            j1, j2)), (mi, -j, j))
    assert JzKetCoupled(j, m, (j1, j2)).rewrite('Jy') == \
        Sum(WignerD(j, mi, m, pi*Rational(3, 2), pi/2, pi/2) * JyKetCoupled(j,
            mi, (j1, j2)), (mi, -j, j))
    assert JxKetCoupled(j, m, (j1, j2)).rewrite('Jz') == \
        Sum(WignerD(j, mi, m, 0, pi/2, 0) * JzKetCoupled(j, mi, (
            j1, j2)), (mi, -j, j))
    assert JyKetCoupled(j, m, (j1, j2)).rewrite('Jz') == \
        Sum(WignerD(j, mi, m, pi*Rational(3, 2), -pi/2, pi/2) * JzKetCoupled(
            j, mi, (j1, j2)), (mi, -j, j))


def test_innerproducts_of_rewritten_states():
    # Numerical
    assert qapply(JxBra(1, 1)*JxKet(1, 1).rewrite('Jy')).doit() == 1
    assert qapply(JxBra(1, 0)*JxKet(1, 0).rewrite('Jy')).doit() == 1
    assert qapply(JxBra(1, -1)*JxKet(1, -1).rewrite('Jy')).doit() == 1
    assert qapply(JxBra(1, 1)*JxKet(1, 1).rewrite('Jz')).doit() == 1
    assert qapply(JxBra(1, 0)*JxKet(1, 0).rewrite('Jz')).doit() == 1
    assert qapply(JxBra(1, -1)*JxKet(1, -1).rewrite('Jz')).doit() == 1
    assert qapply(JyBra(1, 1)*JyKet(1, 1).rewrite('Jx')).doit() == 1
    assert qapply(JyBra(1, 0)*JyKet(1, 0).rewrite('Jx')).doit() == 1
    assert qapply(JyBra(1, -1)*JyKet(1, -1).rewrite('Jx')).doit() == 1
    assert qapply(JyBra(1, 1)*JyKet(1, 1).rewrite('Jz')).doit() == 1
    assert qapply(JyBra(1, 0)*JyKet(1, 0).rewrite('Jz')).doit() == 1
    assert qapply(JyBra(1, -1)*JyKet(1, -1).rewrite('Jz')).doit() == 1
    assert qapply(JyBra(1, 1)*JyKet(1, 1).rewrite('Jz')).doit() == 1
    assert qapply(JyBra(1, 0)*JyKet(1, 0).rewrite('Jz')).doit() == 1
    assert qapply(JyBra(1, -1)*JyKet(1, -1).rewrite('Jz')).doit() == 1
    assert qapply(JzBra(1, 1)*JzKet(1, 1).rewrite('Jy')).doit() == 1
    assert qapply(JzBra(1, 0)*JzKet(1, 0).rewrite('Jy')).doit() == 1
    assert qapply(JzBra(1, -1)*JzKet(1, -1).rewrite('Jy')).doit() == 1
    assert qapply(JxBra(1, 1)*JxKet(1, 0).rewrite('Jy')).doit() == 0
    assert qapply(JxBra(1, 1)*JxKet(1, -1).rewrite('Jy')) == 0
    assert qapply(JxBra(1, 1)*JxKet(1, 0).rewrite('Jz')).doit() == 0
    assert qapply(JxBra(1, 1)*JxKet(1, -1).rewrite('Jz')) == 0
    assert qapply(JyBra(1, 1)*JyKet(1, 0).rewrite('Jx')).doit() == 0
    assert qapply(JyBra(1, 1)*JyKet(1, -1).rewrite('Jx')) == 0
    assert qapply(JyBra(1, 1)*JyKet(1, 0).rewrite('Jz')).doit() == 0
    assert qapply(JyBra(1, 1)*JyKet(1, -1).rewrite('Jz')) == 0
    assert qapply(JzBra(1, 1)*JzKet(1, 0).rewrite('Jx')).doit() == 0
    assert qapply(JzBra(1, 1)*JzKet(1, -1).rewrite('Jx')) == 0
    assert qapply(JzBra(1, 1)*JzKet(1, 0).rewrite('Jy')).doit() == 0
    assert qapply(JzBra(1, 1)*JzKet(1, -1).rewrite('Jy')) == 0
    assert qapply(JxBra(1, 0)*JxKet(1, 1).rewrite('Jy')) == 0
    assert qapply(JxBra(1, 0)*JxKet(1, -1).rewrite('Jy')) == 0
    assert qapply(JxBra(1, 0)*JxKet(1, 1).rewrite('Jz')) == 0
    assert qapply(JxBra(1, 0)*JxKet(1, -1).rewrite('Jz')) == 0
    assert qapply(JyBra(1, 0)*JyKet(1, 1).rewrite('Jx')) == 0
    assert qapply(JyBra(1, 0)*JyKet(1, -1).rewrite('Jx')) == 0
    assert qapply(JyBra(1, 0)*JyKet(1, 1).rewrite('Jz')) == 0
    assert qapply(JyBra(1, 0)*JyKet(1, -1).rewrite('Jz')) == 0
    assert qapply(JzBra(1, 0)*JzKet(1, 1).rewrite('Jx')) == 0
    assert qapply(JzBra(1, 0)*JzKet(1, -1).rewrite('Jx')) == 0
    assert qapply(JzBra(1, 0)*JzKet(1, 1).rewrite('Jy')) == 0
    assert qapply(JzBra(1, 0)*JzKet(1, -1).rewrite('Jy')) == 0
    assert qapply(JxBra(1, -1)*JxKet(1, 1).rewrite('Jy')) == 0
    assert qapply(JxBra(1, -1)*JxKet(1, 0).rewrite('Jy')).doit() == 0
    assert qapply(JxBra(1, -1)*JxKet(1, 1).rewrite('Jz')) == 0
    assert qapply(JxBra(1, -1)*JxKet(1, 0).rewrite('Jz')).doit() == 0
    assert qapply(JyBra(1, -1)*JyKet(1, 1).rewrite('Jx')) == 0
    assert qapply(JyBra(1, -1)*JyKet(1, 0).rewrite('Jx')).doit() == 0
    assert qapply(JyBra(1, -1)*JyKet(1, 1).rewrite('Jz')) == 0
    assert qapply(JyBra(1, -1)*JyKet(1, 0).rewrite('Jz')).doit() == 0
    assert qapply(JzBra(1, -1)*JzKet(1, 1).rewrite('Jx')) == 0
    assert qapply(JzBra(1, -1)*JzKet(1, 0).rewrite('Jx')).doit() == 0
    assert qapply(JzBra(1, -1)*JzKet(1, 1).rewrite('Jy')) == 0
    assert qapply(JzBra(1, -1)*JzKet(1, 0).rewrite('Jy')).doit() == 0


def test_uncouple_2_coupled_states():
    # j1=1/2, j2=1/2
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple(
            TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple(
            TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple(
            TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple(
            TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) )))
    # j1=1/2, j2=1
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1)) == \
        expand(uncouple(
            couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0)) == \
        expand(uncouple(
            couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1)) == \
        expand(uncouple(
            couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1)) == \
        expand(uncouple(
            couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0)) == \
        expand(uncouple(
            couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1)) == \
        expand(uncouple(
            couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1)) )))
    # j1=1, j2=1
    assert TensorProduct(JzKet(1, 1), JzKet(1, 1)) == \
        expand(uncouple(couple( TensorProduct(JzKet(1, 1), JzKet(1, 1)) )))
    assert TensorProduct(JzKet(1, 1), JzKet(1, 0)) == \
        expand(uncouple(couple( TensorProduct(JzKet(1, 1), JzKet(1, 0)) )))
    assert TensorProduct(JzKet(1, 1), JzKet(1, -1)) == \
        expand(uncouple(couple( TensorProduct(JzKet(1, 1), JzKet(1, -1)) )))
    assert TensorProduct(JzKet(1, 0), JzKet(1, 1)) == \
        expand(uncouple(couple( TensorProduct(JzKet(1, 0), JzKet(1, 1)) )))
    assert TensorProduct(JzKet(1, 0), JzKet(1, 0)) == \
        expand(uncouple(couple( TensorProduct(JzKet(1, 0), JzKet(1, 0)) )))
    assert TensorProduct(JzKet(1, 0), JzKet(1, -1)) == \
        expand(uncouple(couple( TensorProduct(JzKet(1, 0), JzKet(1, -1)) )))
    assert TensorProduct(JzKet(1, -1), JzKet(1, 1)) == \
        expand(uncouple(couple( TensorProduct(JzKet(1, -1), JzKet(1, 1)) )))
    assert TensorProduct(JzKet(1, -1), JzKet(1, 0)) == \
        expand(uncouple(couple( TensorProduct(JzKet(1, -1), JzKet(1, 0)) )))
    assert TensorProduct(JzKet(1, -1), JzKet(1, -1)) == \
        expand(uncouple(couple( TensorProduct(JzKet(1, -1), JzKet(1, -1)) )))


def test_uncouple_3_coupled_states():
    # Default coupling
    # j1=1/2, j2=1/2, j3=1/2
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(
            S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S(
            1)/2, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S(
            1)/2, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S(
            1)/2, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.NegativeOne/
               2), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) )))
    # j1=1/2, j2=1, j3=1/2
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(
            JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) )))
    # Coupling j1+j3=j13, j13+j2=j
    # j1=1/2, j2=1/2, j3=1/2
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(
            S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(
            S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(
            S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(
            S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(
            S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(
            S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(
            S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(
            S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) )))
    # j1=1/2, j2=1, j3=1/2
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            1)/2), JzKet(1, 1), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            1)/2), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            1)/2), JzKet(1, 0), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            1)/2), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            1)/2), JzKet(1, -1), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            1)/2), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            -1)/2), JzKet(1, 1), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            -1)/2), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            -1)/2), JzKet(1, 0), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            -1)/2), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S(
            -1)/2), JzKet(1, -1), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.NegativeOne/
               2), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) )))


@slow
def test_uncouple_4_coupled_states():
    # j1=1/2, j2=1/2, j3=1/2, j4=1/2
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(
            S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S(
            1)/2, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S(
            1)/2, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S(
            1)/2, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S(
            1)/2, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S(
            1)/2, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(
            S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S(
            1)/2, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S(
            1)/2, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S(
            1)/2, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S(
            1)/2, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S(
            1)/2, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) )))
    # j1=1/2, j2=1/2, j3=1, j4=1/2
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half),
               JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half),
               JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half),
               JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half),
               JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half),
               JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(
            S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half),
               JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(
            S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half),
               JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(
            S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(
            S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(
            S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)),
               JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)),
               JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)),
               JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)),
               JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)),
               JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(
            S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)),
               JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(
            S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)),
               JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(
            S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(
            S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(
            S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) )))
    # Couple j1+j3=j13, j2+j4=j24, j13+j24=j
    # j1=1/2, j2=1/2, j3=1/2, j4=1/2
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    # j1=1/2, j2=1/2, j3=1, j4=1/2
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, S.Half)), ((1, 3), (2, 4), (1, 2)) )))
    assert TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))) == \
        expand(uncouple(couple( TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (2, 4), (1, 2)) )))


def test_uncouple_2_coupled_states_numerical():
    # j1=1/2, j2=1/2
    assert uncouple(JzKetCoupled(0, 0, (S.Half, S.Half))) == \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))/2 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))/2
    assert uncouple(JzKetCoupled(1, 1, (S.Half, S.Half))) == \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half))
    assert uncouple(JzKetCoupled(1, 0, (S.Half, S.Half))) == \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))/2 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))/2
    assert uncouple(JzKetCoupled(1, -1, (S.Half, S.Half))) == \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))
    # j1=1, j2=1/2
    assert uncouple(JzKetCoupled(S.Half, S.Half, (1, S.Half))) == \
        -sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(S.Half, S.Half))/3 + \
        sqrt(6)*TensorProduct(JzKet(1, 1), JzKet(S.Half, Rational(-1, 2)))/3
    assert uncouple(JzKetCoupled(S.Half, Rational(-1, 2), (1, S.Half))) == \
        sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(S.Half, Rational(-1, 2)))/3 - \
        sqrt(6)*TensorProduct(JzKet(1, -1), JzKet(S.Half, S.Half))/3
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(3, 2), (1, S.Half))) == \
        TensorProduct(JzKet(1, 1), JzKet(S.Half, S.Half))
    assert uncouple(JzKetCoupled(Rational(3, 2), S.Half, (1, S.Half))) == \
        sqrt(3)*TensorProduct(JzKet(1, 1), JzKet(S.Half, Rational(-1, 2)))/3 + \
        sqrt(6)*TensorProduct(JzKet(1, 0), JzKet(S.Half, S.Half))/3
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(-1, 2), (1, S.Half))) == \
        sqrt(6)*TensorProduct(JzKet(1, 0), JzKet(S.Half, Rational(-1, 2)))/3 + \
        sqrt(3)*TensorProduct(JzKet(1, -1), JzKet(S.Half, S.Half))/3
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(-3, 2), (1, S.Half))) == \
        TensorProduct(JzKet(1, -1), JzKet(S.Half, Rational(-1, 2)))
    # j1=1, j2=1
    assert uncouple(JzKetCoupled(0, 0, (1, 1))) == \
        sqrt(3)*TensorProduct(JzKet(1, 1), JzKet(1, -1))/3 - \
        sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(1, 0))/3 + \
        sqrt(3)*TensorProduct(JzKet(1, -1), JzKet(1, 1))/3
    assert uncouple(JzKetCoupled(1, 1, (1, 1))) == \
        sqrt(2)*TensorProduct(JzKet(1, 1), JzKet(1, 0))/2 - \
        sqrt(2)*TensorProduct(JzKet(1, 0), JzKet(1, 1))/2
    assert uncouple(JzKetCoupled(1, 0, (1, 1))) == \
        sqrt(2)*TensorProduct(JzKet(1, 1), JzKet(1, -1))/2 - \
        sqrt(2)*TensorProduct(JzKet(1, -1), JzKet(1, 1))/2
    assert uncouple(JzKetCoupled(1, -1, (1, 1))) == \
        sqrt(2)*TensorProduct(JzKet(1, 0), JzKet(1, -1))/2 - \
        sqrt(2)*TensorProduct(JzKet(1, -1), JzKet(1, 0))/2
    assert uncouple(JzKetCoupled(2, 2, (1, 1))) == \
        TensorProduct(JzKet(1, 1), JzKet(1, 1))
    assert uncouple(JzKetCoupled(2, 1, (1, 1))) == \
        sqrt(2)*TensorProduct(JzKet(1, 1), JzKet(1, 0))/2 + \
        sqrt(2)*TensorProduct(JzKet(1, 0), JzKet(1, 1))/2
    assert uncouple(JzKetCoupled(2, 0, (1, 1))) == \
        sqrt(6)*TensorProduct(JzKet(1, 1), JzKet(1, -1))/6 + \
        sqrt(6)*TensorProduct(JzKet(1, 0), JzKet(1, 0))/3 + \
        sqrt(6)*TensorProduct(JzKet(1, -1), JzKet(1, 1))/6
    assert uncouple(JzKetCoupled(2, -1, (1, 1))) == \
        sqrt(2)*TensorProduct(JzKet(1, 0), JzKet(1, -1))/2 + \
        sqrt(2)*TensorProduct(JzKet(1, -1), JzKet(1, 0))/2
    assert uncouple(JzKetCoupled(2, -2, (1, 1))) == \
        TensorProduct(JzKet(1, -1), JzKet(1, -1))


def test_uncouple_3_coupled_states_numerical():
    # Default coupling
    # j1=1/2, j2=1/2, j3=1/2
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half))) == \
        TensorProduct(JzKet(
            S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half))
    assert uncouple(JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half))) == \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half))/3 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))/3 + \
        sqrt(3)*TensorProduct(JzKet(
            S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))/3
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half))) == \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))/3 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))/3 + \
        sqrt(3)*TensorProduct(JzKet(
            S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))/3
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half))) == \
        TensorProduct(JzKet(
            S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))
    # j1=1/2, j2=1/2, j3=1
    assert uncouple(JzKetCoupled(2, 2, (S.Half, S.Half, 1))) == \
        TensorProduct(
            JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1))
    assert uncouple(JzKetCoupled(2, 1, (S.Half, S.Half, 1))) == \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1))/2 + \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))/2 + \
        sqrt(2)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0))/2
    assert uncouple(JzKetCoupled(2, 0, (S.Half, S.Half, 1))) == \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0))/3 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))/3 + \
        sqrt(6)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1))/6
    assert uncouple(JzKetCoupled(2, -1, (S.Half, S.Half, 1))) == \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))/2 + \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1))/2 + \
        TensorProduct(
            JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(2, -2, (S.Half, S.Half, 1))) == \
        TensorProduct(
            JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1))
    assert uncouple(JzKetCoupled(1, 1, (S.Half, S.Half, 1))) == \
        -TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1))/2 - \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))/2 + \
        sqrt(2)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0))/2
    assert uncouple(JzKetCoupled(1, 0, (S.Half, S.Half, 1))) == \
        -sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))/2 + \
        sqrt(2)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(1, -1, (S.Half, S.Half, 1))) == \
        -sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))/2 + \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1))/2 + \
        TensorProduct(
            JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1))/2
    # j1=1/2, j2=1, j3=1
    assert uncouple(JzKetCoupled(Rational(5, 2), Rational(5, 2), (S.Half, 1, 1))) == \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 1))
    assert uncouple(JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, 1, 1))) == \
        sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/5 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/5 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1),
             JzKet(1, 0))/5
    assert uncouple(JzKetCoupled(Rational(5, 2), S.Half, (S.Half, 1, 1))) == \
        sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/5 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/5 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/10 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/5 + \
        sqrt(10)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/10
    assert uncouple(JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1))) == \
        sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/10 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/5 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/10 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/5 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0),
             JzKet(1, -1))/5
    assert uncouple(JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, 1, 1))) == \
        sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/5 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/5 + \
        sqrt(5)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, -1))/5
    assert uncouple(JzKetCoupled(Rational(5, 2), Rational(-5, 2), (S.Half, 1, 1))) == \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, -1))
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1))) == \
        -sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/15 - \
        2*sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1),
             JzKet(1, 0))/5
    assert uncouple(JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1))) == \
        -4*sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/15 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/15 - \
        2*sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/15 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1),
             JzKet(1, -1))/5
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1))) == \
        -sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/5 - \
        sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/15 + \
        2*sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/15 - \
        sqrt(5)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/15 + \
        4*sqrt(5)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/15
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1))) == \
        -sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/5 + \
        2*sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/15 + \
        sqrt(30)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, -1))/15
    assert uncouple(JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1))) == \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/3 - \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/3 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/6 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/3 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1),
             JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1))) == \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/2 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/3 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/6 - \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/3 + \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/3
    # j1=1, j2=1, j3=1
    assert uncouple(JzKetCoupled(3, 3, (1, 1, 1))) == \
        TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, 1))
    assert uncouple(JzKetCoupled(3, 2, (1, 1, 1))) == \
        sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, 1))/3 + \
        sqrt(3)*TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 1))/3 + \
        sqrt(3)*TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, 0))/3
    assert uncouple(JzKetCoupled(3, 1, (1, 1, 1))) == \
        sqrt(15)*TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 1))/15 + \
        2*sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 1))/15 + \
        2*sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 1))/15 + \
        2*sqrt(15)*TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, -1))/15
    assert uncouple(JzKetCoupled(3, 0, (1, 1, 1))) == \
        sqrt(10)*TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 1))/10 + \
        sqrt(10)*TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 0))/10 + \
        sqrt(10)*TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 1))/10 + \
        sqrt(10)*TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 0))/5 + \
        sqrt(10)*TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, -1))/10 + \
        sqrt(10)*TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 0))/10 + \
        sqrt(10)*TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, -1))/10
    assert uncouple(JzKetCoupled(3, -1, (1, 1, 1))) == \
        sqrt(15)*TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 1))/15 + \
        2*sqrt(15)*TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, -1))/15 + \
        2*sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 0))/15 + \
        2*sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, -1))/15 + \
        sqrt(15)*TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, -1))/15
    assert uncouple(JzKetCoupled(3, -2, (1, 1, 1))) == \
        sqrt(3)*TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 0))/3 + \
        sqrt(3)*TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, -1))/3 + \
        sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, -1))/3
    assert uncouple(JzKetCoupled(3, -3, (1, 1, 1))) == \
        TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, -1))
    assert uncouple(JzKetCoupled(2, 2, (1, 1, 1))) == \
        -sqrt(6)*TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, 1))/6 - \
        sqrt(6)*TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 1))/6 + \
        sqrt(6)*TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, 0))/3
    assert uncouple(JzKetCoupled(2, 1, (1, 1, 1))) == \
        -sqrt(3)*TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 1))/6 - \
        sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 1))/3 + \
        sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, 0))/6 - \
        sqrt(3)*TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 1))/6 + \
        sqrt(3)*TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 0))/6 + \
        sqrt(3)*TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, -1))/3
    assert uncouple(JzKetCoupled(2, 0, (1, 1, 1))) == \
        -TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 1))/2 - \
        TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 1))/2 + \
        TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, -1))/2 + \
        TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(2, -1, (1, 1, 1))) == \
        -sqrt(3)*TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 1))/3 - \
        sqrt(3)*TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 0))/6 + \
        sqrt(3)*TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, -1))/6 - \
        sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 0))/6 + \
        sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, -1))/3 + \
        sqrt(3)*TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, -1))/6
    assert uncouple(JzKetCoupled(2, -2, (1, 1, 1))) == \
        -sqrt(6)*TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 0))/3 + \
        sqrt(6)*TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, -1))/6 + \
        sqrt(6)*TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, -1))/6
    assert uncouple(JzKetCoupled(1, 1, (1, 1, 1))) == \
        sqrt(15)*TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 1))/30 + \
        sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 1))/15 - \
        sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, 0))/10 + \
        sqrt(15)*TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 1))/30 - \
        sqrt(15)*TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 0))/10 + \
        sqrt(15)*TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, -1))/5
    assert uncouple(JzKetCoupled(1, 0, (1, 1, 1))) == \
        sqrt(15)*TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 1))/10 - \
        sqrt(15)*TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 1))/10 - \
        2*sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, -1))/10 - \
        sqrt(15)*TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, -1))/10
    assert uncouple(JzKetCoupled(1, -1, (1, 1, 1))) == \
        sqrt(15)*TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 1))/5 - \
        sqrt(15)*TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 0))/10 + \
        sqrt(15)*TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, -1))/30 - \
        sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 0))/10 + \
        sqrt(15)*TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, -1))/15 + \
        sqrt(15)*TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, -1))/30
    # Defined j13
    # j1=1/2, j2=1/2, j3=1, j13=1/2
    assert uncouple(JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )) == \
        -sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1))/3 + \
        sqrt(3)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0))/3
    assert uncouple(JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )) == \
        -sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))/3 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0))/6 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))/6 + \
        sqrt(3)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1))/3
    assert uncouple(JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )) == \
        -sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))/3 + \
        sqrt(6)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1))/3
    # j1=1/2, j2=1, j3=1, j13=1/2
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))))) == \
        -sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/3 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1),
             JzKet(1, 0))/3
    assert uncouple(JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))))) == \
        -2*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/3 - \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/3 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/3 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1),
             JzKet(1, -1))/3
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))))) == \
        -sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/3 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/3 + \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/3 + \
        2*TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/3
    assert uncouple(JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))))) == \
        -sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/3 + \
        sqrt(6)*TensorProduct(
            JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, -1))/3
    # j1=1, j2=1, j3=1, j13=1
    assert uncouple(JzKetCoupled(2, 2, (1, 1, 1), ((1, 3, 1), (1, 2, 2)))) == \
        -sqrt(2)*TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, 1))/2 + \
        sqrt(2)*TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, 0))/2
    assert uncouple(JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 1), (1, 2, 2)))) == \
        -TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 1))/2 - \
        TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 1))/2 + \
        TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 0))/2 + \
        TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 2)))) == \
        -sqrt(3)*TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 1))/3 - \
        sqrt(3)*TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 0))/6 - \
        sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 1))/6 + \
        sqrt(3)*TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, -1))/6 + \
        sqrt(3)*TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 0))/6 + \
        sqrt(3)*TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, -1))/3
    assert uncouple(JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 1), (1, 2, 2)))) == \
        -TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 1))/2 - \
        TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 0))/2 + \
        TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, -1))/2 + \
        TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(2, -2, (1, 1, 1), ((1, 3, 1), (1, 2, 2)))) == \
        -sqrt(2)*TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 0))/2 + \
        sqrt(2)*TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 1), (1, 2, 1)))) == \
        TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 1))/2 - \
        TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 1))/2 + \
        TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 0))/2 - \
        TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 1)))) == \
        TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 0))/2 - \
        TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 1))/2 - \
        TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, -1))/2 + \
        TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 0))/2
    assert uncouple(JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 1), (1, 2, 1)))) == \
        -TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 1))/2 + \
        TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 0))/2 - \
        TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, -1))/2 + \
        TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, -1))/2


def test_uncouple_4_coupled_states_numerical():
    # j1=1/2, j2=1/2, j3=1, j4=1, default coupling
    assert uncouple(JzKetCoupled(3, 3, (S.Half, S.Half, 1, 1))) == \
        TensorProduct(JzKet(
            S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 1))
    assert uncouple(JzKetCoupled(3, 2, (S.Half, S.Half, 1, 1))) == \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/3 + \
        sqrt(3)*TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/3
    assert uncouple(JzKetCoupled(3, 1, (S.Half, S.Half, 1, 1))) == \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/15 + \
        2*sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, 1), JzKet(1, -1))/15
    assert uncouple(JzKetCoupled(3, 0, (S.Half, S.Half, 1, 1))) == \
        sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/10 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/10 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/10 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/5 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/10 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/10 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/5 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/10 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/10 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, 0), JzKet(1, -1))/10
    assert uncouple(JzKetCoupled(3, -1, (S.Half, S.Half, 1, 1))) == \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/15 + \
        2*sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, -1), JzKet(1, -1))/15
    assert uncouple(JzKetCoupled(3, -2, (S.Half, S.Half, 1, 1))) == \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/3 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/3 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, -1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             Rational(-1, 2)), JzKet(1, -1), JzKet(1, -1))/6
    assert uncouple(JzKetCoupled(3, -3, (S.Half, S.Half, 1, 1))) == \
        TensorProduct(JzKet(S.Half, -S(
            1)/2), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, -1))
    assert uncouple(JzKetCoupled(2, 2, (S.Half, S.Half, 1, 1))) == \
        -sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 1))/6 - \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/6 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/3
    assert uncouple(JzKetCoupled(2, 1, (S.Half, S.Half, 1, 1))) == \
        -sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/6 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/12 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/12 - \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/6 + \
        sqrt(3)*TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/3
    assert uncouple(JzKetCoupled(2, 0, (S.Half, S.Half, 1, 1))) == \
        -TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/2 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/4 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/4 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/4 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/4 + \
        TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(2, -1, (S.Half, S.Half, 1, 1))) == \
        -sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/3 - \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/6 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/12 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/6 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/12 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, -1), JzKet(1, -1))/6
    assert uncouple(JzKetCoupled(2, -2, (S.Half, S.Half, 1, 1))) == \
        -sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/3 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, -1))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             Rational(-1, 2)), JzKet(1, -1), JzKet(1, -1))/6
    assert uncouple(JzKetCoupled(1, 1, (S.Half, S.Half, 1, 1))) == \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/30 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/30 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/20 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/30 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/20 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/30 - \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/10 + \
        sqrt(15)*TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/5
    assert uncouple(JzKetCoupled(1, 0, (S.Half, S.Half, 1, 1))) == \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/10 - \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/20 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/20 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/20 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/20 - \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, 0), JzKet(1, -1))/10
    assert uncouple(JzKetCoupled(1, -1, (S.Half, S.Half, 1, 1))) == \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/5 - \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/10 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/30 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/20 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/30 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/20 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/30 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, -1), JzKet(1, -1))/30
    # j1=1/2, j2=1/2, j3=1, j4=1, j12=1, j34=1
    assert uncouple(JzKetCoupled(2, 2, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 1), (1, 3, 2)))) == \
        -sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/2 + \
        sqrt(2)*TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/2
    assert uncouple(JzKetCoupled(2, 1, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 1), (1, 3, 2)))) == \
        -sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/4 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/4 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/4 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/4 - \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/2 + \
        TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(2, 0, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 1), (1, 3, 2)))) == \
        -sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/6 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/6 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/6 - \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/6 + \
        sqrt(3)*TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/6
    assert uncouple(JzKetCoupled(2, -1, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 1), (1, 3, 2)))) == \
        -TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/2 + \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/2 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/4 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/4 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/4 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/4
    assert uncouple(JzKetCoupled(2, -2, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 1), (1, 3, 2)))) == \
        -sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/2 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half,
             Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(1, 1, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 1), (1, 3, 1)))) == \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/4 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/4 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/4 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/4 - \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/2 + \
        TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(1, 0, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 1), (1, 3, 1)))) == \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/2 - \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/2 - \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/2 + \
        TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(1, -1, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 1), (1, 3, 1)))) == \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/2 - \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/2 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/4 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/4 - \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/4 + \
        sqrt(2)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/4
    # j1=1/2, j2=1/2, j3=1, j4=1, j12=1, j34=2
    assert uncouple(JzKetCoupled(3, 3, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 3)))) == \
        TensorProduct(JzKet(
            S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 1))
    assert uncouple(JzKetCoupled(3, 2, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 3)))) == \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/3 + \
        sqrt(3)*TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/3
    assert uncouple(JzKetCoupled(3, 1, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 3)))) == \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/15 + \
        2*sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, 1), JzKet(1, -1))/15
    assert uncouple(JzKetCoupled(3, 0, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 3)))) == \
        sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/10 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/10 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/10 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/5 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/10 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/10 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/5 + \
        sqrt(5)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/10 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/10 + \
        sqrt(10)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, 0), JzKet(1, -1))/10
    assert uncouple(JzKetCoupled(3, -1, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 3)))) == \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/15 + \
        2*sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/15 + \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, -1), JzKet(1, -1))/15
    assert uncouple(JzKetCoupled(3, -2, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 3)))) == \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/3 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/3 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, -1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             Rational(-1, 2)), JzKet(1, -1), JzKet(1, -1))/6
    assert uncouple(JzKetCoupled(3, -3, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 3)))) == \
        TensorProduct(JzKet(S.Half, -S(
            1)/2), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, -1))
    assert uncouple(JzKetCoupled(2, 2, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 2)))) == \
        -sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 1))/3 - \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/3 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/6
    assert uncouple(JzKetCoupled(2, 1, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 2)))) == \
        -sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/3 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/12 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/12 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/12 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/12 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/3 + \
        sqrt(3)*TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/6
    assert uncouple(JzKetCoupled(2, 0, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 2)))) == \
        -TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/2 - \
        TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/2 + \
        TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/2 + \
        TensorProduct(JzKet(S(
            1)/2, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/2
    assert uncouple(JzKetCoupled(2, -1, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 2)))) == \
        -sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/6 - \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/3 - \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/6 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/12 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/12 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/12 + \
        sqrt(6)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/12 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, -1), JzKet(1, -1))/3
    assert uncouple(JzKetCoupled(2, -2, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 2)))) == \
        -sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/6 - \
        sqrt(6)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/6 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, -1))/3 + \
        sqrt(3)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             Rational(-1, 2)), JzKet(1, -1), JzKet(1, -1))/3
    assert uncouple(JzKetCoupled(1, 1, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 1)))) == \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))/5 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))/20 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))/20 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/20 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/20 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/30 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, 1), JzKet(1, -1))/30
    assert uncouple(JzKetCoupled(1, 0, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 1)))) == \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))/10 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))/10 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))/30 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))/15 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))/30 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/30 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/15 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/30 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/10 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, 0), JzKet(1, -1))/10
    assert uncouple(JzKetCoupled(1, -1, (S.Half, S.Half, 1, 1), ((1, 2, 1), (3, 4, 2), (1, 3, 1)))) == \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))/30 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))/15 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))/30 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))/20 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))/20 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))/20 - \
        sqrt(30)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))/20 + \
        sqrt(15)*TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half,
             S.Half), JzKet(1, -1), JzKet(1, -1))/5


def test_uncouple_symbolic():
    assert uncouple(JzKetCoupled(j, m, (j1, j2) )) == \
        Sum(CG(j1, m1, j2, m2, j, m) *
            TensorProduct(JzKet(j1, m1), JzKet(j2, m2)),
            (m1, -j1, j1), (m2, -j2, j2))
    assert uncouple(JzKetCoupled(j, m, (j1, j2, j3) )) == \
        Sum(CG(j1, m1, j2, m2, j1 + j2, m1 + m2) * CG(j1 + j2, m1 + m2, j3, m3, j, m) *
            TensorProduct(JzKet(j1, m1), JzKet(j2, m2), JzKet(j3, m3)),
            (m1, -j1, j1), (m2, -j2, j2), (m3, -j3, j3))
    assert uncouple(JzKetCoupled(j, m, (j1, j2, j3), ((1, 3, j13), (1, 2, j)) )) == \
        Sum(CG(j1, m1, j3, m3, j13, m1 + m3) * CG(j13, m1 + m3, j2, m2, j, m) *
            TensorProduct(JzKet(j1, m1), JzKet(j2, m2), JzKet(j3, m3)),
            (m1, -j1, j1), (m2, -j2, j2), (m3, -j3, j3))
    assert uncouple(JzKetCoupled(j, m, (j1, j2, j3, j4) )) == \
        Sum(CG(j1, m1, j2, m2, j1 + j2, m1 + m2) * CG(j1 + j2, m1 + m2, j3, m3, j1 + j2 + j3, m1 + m2 + m3) * CG(j1 + j2 + j3, m1 + m2 + m3, j4, m4, j, m) *
            TensorProduct(
                JzKet(j1, m1), JzKet(j2, m2), JzKet(j3, m3), JzKet(j4, m4)),
            (m1, -j1, j1), (m2, -j2, j2), (m3, -j3, j3), (m4, -j4, j4))
    assert uncouple(JzKetCoupled(j, m, (j1, j2, j3, j4), ((1, 3, j13), (2, 4, j24), (1, 2, j)) )) ==  \
        Sum(CG(j1, m1, j3, m3, j13, m1 + m3) * CG(j2, m2, j4, m4, j24, m2 + m4) * CG(j13, m1 + m3, j24, m2 + m4, j, m) *
            TensorProduct(
                JzKet(j1, m1), JzKet(j2, m2), JzKet(j3, m3), JzKet(j4, m4)),
            (m1, -j1, j1), (m2, -j2, j2), (m3, -j3, j3), (m4, -j4, j4))


def test_couple_2_states():
    # j1=1/2, j2=1/2
    assert JzKetCoupled(0, 0, (S.Half, S.Half)) == \
        expand(couple(uncouple( JzKetCoupled(0, 0, (S.Half, S.Half)) )))
    assert JzKetCoupled(1, 1, (S.Half, S.Half)) == \
        expand(couple(uncouple( JzKetCoupled(1, 1, (S.Half, S.Half)) )))
    assert JzKetCoupled(1, 0, (S.Half, S.Half)) == \
        expand(couple(uncouple( JzKetCoupled(1, 0, (S.Half, S.Half)) )))
    assert JzKetCoupled(1, -1, (S.Half, S.Half)) == \
        expand(couple(uncouple( JzKetCoupled(1, -1, (S.Half, S.Half)) )))
    # j1=1, j2=1/2
    assert JzKetCoupled(S.Half, S.Half, (1, S.Half)) == \
        expand(couple(uncouple( JzKetCoupled(S.Half, S.Half, (1, S.Half)) )))
    assert JzKetCoupled(S.Half, Rational(-1, 2), (1, S.Half)) == \
        expand(couple(uncouple( JzKetCoupled(S.Half, Rational(-1, 2), (1, S.Half)) )))
    assert JzKetCoupled(Rational(3, 2), Rational(3, 2), (1, S.Half)) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), Rational(3, 2), (1, S.Half)) )))
    assert JzKetCoupled(Rational(3, 2), S.Half, (1, S.Half)) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), S.Half, (1, S.Half)) )))
    assert JzKetCoupled(Rational(3, 2), Rational(-1, 2), (1, S.Half)) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), Rational(-1, 2), (1, S.Half)) )))
    assert JzKetCoupled(Rational(3, 2), Rational(-3, 2), (1, S.Half)) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), Rational(-3, 2), (1, S.Half)) )))
    # j1=1, j2=1
    assert JzKetCoupled(0, 0, (1, 1)) == \
        expand(couple(uncouple( JzKetCoupled(0, 0, (1, 1)) )))
    assert JzKetCoupled(1, 1, (1, 1)) == \
        expand(couple(uncouple( JzKetCoupled(1, 1, (1, 1)) )))
    assert JzKetCoupled(1, 0, (1, 1)) == \
        expand(couple(uncouple( JzKetCoupled(1, 0, (1, 1)) )))
    assert JzKetCoupled(1, -1, (1, 1)) == \
        expand(couple(uncouple( JzKetCoupled(1, -1, (1, 1)) )))
    assert JzKetCoupled(2, 2, (1, 1)) == \
        expand(couple(uncouple( JzKetCoupled(2, 2, (1, 1)) )))
    assert JzKetCoupled(2, 1, (1, 1)) == \
        expand(couple(uncouple( JzKetCoupled(2, 1, (1, 1)) )))
    assert JzKetCoupled(2, 0, (1, 1)) == \
        expand(couple(uncouple( JzKetCoupled(2, 0, (1, 1)) )))
    assert JzKetCoupled(2, -1, (1, 1)) == \
        expand(couple(uncouple( JzKetCoupled(2, -1, (1, 1)) )))
    assert JzKetCoupled(2, -2, (1, 1)) == \
        expand(couple(uncouple( JzKetCoupled(2, -2, (1, 1)) )))
    # j1=1/2, j2=3/2
    assert JzKetCoupled(1, 1, (S.Half, Rational(3, 2))) == \
        expand(couple(uncouple( JzKetCoupled(1, 1, (S.Half, Rational(3, 2))) )))
    assert JzKetCoupled(1, 0, (S.Half, Rational(3, 2))) == \
        expand(couple(uncouple( JzKetCoupled(1, 0, (S.Half, Rational(3, 2))) )))
    assert JzKetCoupled(1, -1, (S.Half, Rational(3, 2))) == \
        expand(couple(uncouple( JzKetCoupled(1, -1, (S.Half, Rational(3, 2))) )))
    assert JzKetCoupled(2, 2, (S.Half, Rational(3, 2))) == \
        expand(couple(uncouple( JzKetCoupled(2, 2, (S.Half, Rational(3, 2))) )))
    assert JzKetCoupled(2, 1, (S.Half, Rational(3, 2))) == \
        expand(couple(uncouple( JzKetCoupled(2, 1, (S.Half, Rational(3, 2))) )))
    assert JzKetCoupled(2, 0, (S.Half, Rational(3, 2))) == \
        expand(couple(uncouple( JzKetCoupled(2, 0, (S.Half, Rational(3, 2))) )))
    assert JzKetCoupled(2, -1, (S.Half, Rational(3, 2))) == \
        expand(couple(uncouple( JzKetCoupled(2, -1, (S.Half, Rational(3, 2))) )))
    assert JzKetCoupled(2, -2, (S.Half, Rational(3, 2))) == \
        expand(couple(uncouple( JzKetCoupled(2, -2, (S.Half, Rational(3, 2))) )))


def test_couple_3_states():
    # Default coupling
    # j1=1/2, j2=1/2, j3=1/2
    assert JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half)) == \
        expand(couple(uncouple(
            JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half)) == \
        expand(couple(uncouple(
            JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half)) )))
    # j1=1/2, j2=1/2, j3=1
    assert JzKetCoupled(0, 0, (S.Half, S.Half, 1)) == \
        expand(couple(uncouple( JzKetCoupled(0, 0, (S.Half, S.Half, 1)) )))
    assert JzKetCoupled(1, 1, (S.Half, S.Half, 1)) == \
        expand(couple(uncouple( JzKetCoupled(1, 1, (S.Half, S.Half, 1)) )))
    assert JzKetCoupled(1, 0, (S.Half, S.Half, 1)) == \
        expand(couple(uncouple( JzKetCoupled(1, 0, (S.Half, S.Half, 1)) )))
    assert JzKetCoupled(1, -1, (S.Half, S.Half, 1)) == \
        expand(couple(uncouple( JzKetCoupled(1, -1, (S.Half, S.Half, 1)) )))
    assert JzKetCoupled(2, 2, (S.Half, S.Half, 1)) == \
        expand(couple(uncouple( JzKetCoupled(2, 2, (S.Half, S.Half, 1)) )))
    assert JzKetCoupled(2, 1, (S.Half, S.Half, 1)) == \
        expand(couple(uncouple( JzKetCoupled(2, 1, (S.Half, S.Half, 1)) )))
    assert JzKetCoupled(2, 0, (S.Half, S.Half, 1)) == \
        expand(couple(uncouple( JzKetCoupled(2, 0, (S.Half, S.Half, 1)) )))
    assert JzKetCoupled(2, -1, (S.Half, S.Half, 1)) == \
        expand(couple(uncouple( JzKetCoupled(2, -1, (S.Half, S.Half, 1)) )))
    assert JzKetCoupled(2, -2, (S.Half, S.Half, 1)) == \
        expand(couple(uncouple( JzKetCoupled(2, -2, (S.Half, S.Half, 1)) )))
    # Couple j1+j3=j13, j13+j2=j
    # j1=1/2, j2=1/2, j3=1/2, j13=0
    assert JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 3, 0), (1, 2, S.Half))) == \
        expand(couple(uncouple( JzKetCoupled(S.Half, S.Half, (S.Half, S(
            1)/2, S.Half), ((1, 3, 0), (1, 2, S.Half))) ), ((1, 3), (1, 2)) ))
    assert JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 3, 0), (1, 2, S.Half))) == \
        expand(couple(uncouple( JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S(
            1)/2, S.Half), ((1, 3, 0), (1, 2, S.Half))) ), ((1, 3), (1, 2)) ))
    # j1=1, j2=1/2, j3=1, j13=1
    assert JzKetCoupled(S.Half, S.Half, (1, S.Half, 1), ((1, 3, 1), (1, 2, S.Half))) == \
        expand(couple(uncouple( JzKetCoupled(S.Half, S.Half, (
            1, S.Half, 1), ((1, 3, 1), (1, 2, S.Half))) ), ((1, 3), (1, 2)) ))
    assert JzKetCoupled(S.Half, Rational(-1, 2), (1, S.Half, 1), ((1, 3, 1), (1, 2, S.Half))) == \
        expand(couple(uncouple( JzKetCoupled(S.Half, Rational(-1, 2), (
            1, S.Half, 1), ((1, 3, 1), (1, 2, S.Half))) ), ((1, 3), (1, 2)) ))
    assert JzKetCoupled(Rational(3, 2), Rational(3, 2), (1, S.Half, 1), ((1, 3, 1), (1, 2, Rational(3, 2)))) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), Rational(3, 2), (
            1, S.Half, 1), ((1, 3, 1), (1, 2, Rational(3, 2)))) ), ((1, 3), (1, 2)) ))
    assert JzKetCoupled(Rational(3, 2), S.Half, (1, S.Half, 1), ((1, 3, 1), (1, 2, Rational(3, 2)))) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), S.Half, (
            1, S.Half, 1), ((1, 3, 1), (1, 2, Rational(3, 2)))) ), ((1, 3), (1, 2)) ))
    assert JzKetCoupled(Rational(3, 2), Rational(-1, 2), (1, S.Half, 1), ((1, 3, 1), (1, 2, Rational(3, 2)))) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), Rational(-1, 2), (
            1, S.Half, 1), ((1, 3, 1), (1, 2, Rational(3, 2)))) ), ((1, 3), (1, 2)) ))
    assert JzKetCoupled(Rational(3, 2), Rational(-3, 2), (1, S.Half, 1), ((1, 3, 1), (1, 2, Rational(3, 2)))) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), Rational(-3, 2), (
            1, S.Half, 1), ((1, 3, 1), (1, 2, Rational(3, 2)))) ), ((1, 3), (1, 2)) ))


def test_couple_4_states():
    # Default coupling
    # j1=1/2, j2=1/2, j3=1/2, j4=1/2
    assert JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half)) == \
        expand(couple(
            uncouple( JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half)) == \
        expand(couple(
            uncouple( JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half)) == \
        expand(couple(uncouple(
            JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(2, 2, (S.Half, S.Half, S.Half, S.Half)) == \
        expand(couple(
            uncouple( JzKetCoupled(2, 2, (S.Half, S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(2, 1, (S.Half, S.Half, S.Half, S.Half)) == \
        expand(couple(
            uncouple( JzKetCoupled(2, 1, (S.Half, S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(2, 0, (S.Half, S.Half, S.Half, S.Half)) == \
        expand(couple(
            uncouple( JzKetCoupled(2, 0, (S.Half, S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(2, -1, (S.Half, S.Half, S.Half, S.Half)) == \
        expand(couple(uncouple(
            JzKetCoupled(2, -1, (S.Half, S.Half, S.Half, S.Half)) )))
    assert JzKetCoupled(2, -2, (S.Half, S.Half, S.Half, S.Half)) == \
        expand(couple(uncouple(
            JzKetCoupled(2, -2, (S.Half, S.Half, S.Half, S.Half)) )))
    # j1=1/2, j2=1/2, j3=1/2, j4=1
    assert JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(Rational(5, 2), Rational(5, 2), (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(5, 2), Rational(5, 2), (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1)) )))
    assert JzKetCoupled(Rational(5, 2), Rational(-5, 2), (S.Half, S.Half, S.Half, 1)) == \
        expand(couple(uncouple(
            JzKetCoupled(Rational(5, 2), Rational(-5, 2), (S.Half, S.Half, S.Half, 1)) )))
    # Coupling j1+j3=j13, j2+j4=j24, j13+j24=j
    # j1=1/2, j2=1/2, j3=1/2, j4=1/2, j13=1, j24=0
    assert JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 3, 1), (2, 4, 0), (1, 2, 1)) ) == \
        expand(couple(uncouple( JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 3, 1), (2, 4, 0), (1, 2, 1)) ) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 3, 1), (2, 4, 0), (1, 2, 1)) ) == \
        expand(couple(uncouple( JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 3, 1), (2, 4, 0), (1, 2, 1)) ) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 3, 1), (2, 4, 0), (1, 2, 1)) ) == \
        expand(couple(uncouple( JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 3, 1), (2, 4, 0), (1, 2, 1)) ) ), ((1, 3), (2, 4), (1, 2)) ))
    # j1=1/2, j2=1/2, j3=1/2, j4=1, j13=1, j24=1/2
    assert JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, S.Half)) ) == \
        expand(couple(uncouple( JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, S.Half)) )), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, S.Half)) ) == \
        expand(couple(uncouple( JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, S.Half)) ) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, Rational(3, 2))) ) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, Rational(3, 2))) ) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, Rational(3, 2))) ) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, Rational(3, 2))) ) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, Rational(3, 2))) ) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, Rational(3, 2))) ) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, Rational(3, 2))) ) == \
        expand(couple(uncouple( JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 3, 1), (2, 4, S.Half), (1, 2, Rational(3, 2))) ) ), ((1, 3), (2, 4), (1, 2)) ))
    # j1=1/2, j2=1, j3=1/2, j4=1, j13=0, j24=1
    assert JzKetCoupled(1, 1, (S.Half, 1, S.Half, 1), ((1, 3, 0), (2, 4, 1), (1, 2, 1)) ) == \
        expand(couple(uncouple( JzKetCoupled(1, 1, (S.Half, 1, S.Half, 1), (
            (1, 3, 0), (2, 4, 1), (1, 2, 1))) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(1, 0, (S.Half, 1, S.Half, 1), ((1, 3, 0), (2, 4, 1), (1, 2, 1)) ) == \
        expand(couple(uncouple( JzKetCoupled(1, 0, (S.Half, 1, S.Half, 1), (
            (1, 3, 0), (2, 4, 1), (1, 2, 1))) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(1, -1, (S.Half, 1, S.Half, 1), ((1, 3, 0), (2, 4, 1), (1, 2, 1)) ) == \
        expand(couple(uncouple( JzKetCoupled(1, -1, (S.Half, 1, S.Half, 1), (
            (1, 3, 0), (2, 4, 1), (1, 2, 1))) ), ((1, 3), (2, 4), (1, 2)) ))
    # j1=1/2, j2=1, j3=1/2, j4=1, j13=1, j24=1
    assert JzKetCoupled(0, 0, (S.Half, 1, S.Half, 1), ((1, 3, 1), (2, 4, 1), (1, 2, 0)) ) == \
        expand(couple(uncouple( JzKetCoupled(0, 0, (S.Half, 1, S.Half, 1), (
            (1, 3, 1), (2, 4, 1), (1, 2, 0))) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(1, 1, (S.Half, 1, S.Half, 1), ((1, 3, 1), (2, 4, 1), (1, 2, 1)) ) == \
        expand(couple(uncouple( JzKetCoupled(1, 1, (S.Half, 1, S.Half, 1), (
            (1, 3, 1), (2, 4, 1), (1, 2, 1))) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(1, 0, (S.Half, 1, S.Half, 1), ((1, 3, 1), (2, 4, 1), (1, 2, 1)) ) == \
        expand(couple(uncouple( JzKetCoupled(1, 0, (S.Half, 1, S.Half, 1), (
            (1, 3, 1), (2, 4, 1), (1, 2, 1))) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(1, -1, (S.Half, 1, S.Half, 1), ((1, 3, 1), (2, 4, 1), (1, 2, 1)) ) == \
        expand(couple(uncouple( JzKetCoupled(1, -1, (S.Half, 1, S.Half, 1), (
            (1, 3, 1), (2, 4, 1), (1, 2, 1))) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(2, 2, (S.Half, 1, S.Half, 1), ((1, 3, 1), (2, 4, 1), (1, 2, 2)) ) == \
        expand(couple(uncouple( JzKetCoupled(2, 2, (S.Half, 1, S.Half, 1), (
            (1, 3, 1), (2, 4, 1), (1, 2, 2))) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(2, 1, (S.Half, 1, S.Half, 1), ((1, 3, 1), (2, 4, 1), (1, 2, 2)) ) == \
        expand(couple(uncouple( JzKetCoupled(2, 1, (S.Half, 1, S.Half, 1), (
            (1, 3, 1), (2, 4, 1), (1, 2, 2))) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(2, 0, (S.Half, 1, S.Half, 1), ((1, 3, 1), (2, 4, 1), (1, 2, 2)) ) == \
        expand(couple(uncouple( JzKetCoupled(2, 0, (S.Half, 1, S.Half, 1), (
            (1, 3, 1), (2, 4, 1), (1, 2, 2))) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(2, -1, (S.Half, 1, S.Half, 1), ((1, 3, 1), (2, 4, 1), (1, 2, 2)) ) == \
        expand(couple(uncouple( JzKetCoupled(2, -1, (S.Half, 1, S.Half, 1), (
            (1, 3, 1), (2, 4, 1), (1, 2, 2))) ), ((1, 3), (2, 4), (1, 2)) ))
    assert JzKetCoupled(2, -2, (S.Half, 1, S.Half, 1), ((1, 3, 1), (2, 4, 1), (1, 2, 2)) ) == \
        expand(couple(uncouple( JzKetCoupled(2, -2, (S.Half, 1, S.Half, 1), (
            (1, 3, 1), (2, 4, 1), (1, 2, 2))) ), ((1, 3), (2, 4), (1, 2)) ))


def test_couple_2_states_numerical():
    # j1=1/2, j2=1/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half))) == \
        JzKetCoupled(1, 1, (S.Half, S.Half))
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))) == \
        sqrt(2)*JzKetCoupled(0, 0, (S(
            1)/2, S.Half))/2 + sqrt(2)*JzKetCoupled(1, 0, (S.Half, S.Half))/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))) == \
        -sqrt(2)*JzKetCoupled(0, 0, (S(
            1)/2, S.Half))/2 + sqrt(2)*JzKetCoupled(1, 0, (S.Half, S.Half))/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))) == \
        JzKetCoupled(1, -1, (S.Half, S.Half))
    # j1=1, j2=1/2
    assert couple(TensorProduct(JzKet(1, 1), JzKet(S.Half, S.Half))) == \
        JzKetCoupled(Rational(3, 2), Rational(3, 2), (1, S.Half))
    assert couple(TensorProduct(JzKet(1, 1), JzKet(S.Half, Rational(-1, 2)))) == \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (1, S.Half))/3 + sqrt(
            3)*JzKetCoupled(Rational(3, 2), S.Half, (1, S.Half))/3
    assert couple(TensorProduct(JzKet(1, 0), JzKet(S.Half, S.Half))) == \
        -sqrt(3)*JzKetCoupled(S.Half, S.Half, (1, S.Half))/3 + \
        sqrt(6)*JzKetCoupled(Rational(3, 2), S.Half, (1, S.Half))/3
    assert couple(TensorProduct(JzKet(1, 0), JzKet(S.Half, Rational(-1, 2)))) == \
        sqrt(3)*JzKetCoupled(S.Half, Rational(-1, 2), (1, S.Half))/3 + \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (1, S.Half))/3
    assert couple(TensorProduct(JzKet(1, -1), JzKet(S.Half, S.Half))) == \
        -sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (1, S(
            1)/2))/3 + sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (1, S.Half))/3
    assert couple(TensorProduct(JzKet(1, -1), JzKet(S.Half, Rational(-1, 2)))) == \
        JzKetCoupled(Rational(3, 2), Rational(-3, 2), (1, S.Half))
    # j1=1, j2=1
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 1))) == \
        JzKetCoupled(2, 2, (1, 1))
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 0))) == \
        sqrt(2)*JzKetCoupled(
            1, 1, (1, 1))/2 + sqrt(2)*JzKetCoupled(2, 1, (1, 1))/2
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        sqrt(3)*JzKetCoupled(0, 0, (1, 1))/3 + sqrt(2)*JzKetCoupled(
            1, 0, (1, 1))/2 + sqrt(6)*JzKetCoupled(2, 0, (1, 1))/6
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 1))) == \
        -sqrt(2)*JzKetCoupled(
            1, 1, (1, 1))/2 + sqrt(2)*JzKetCoupled(2, 1, (1, 1))/2
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 0))) == \
        -sqrt(3)*JzKetCoupled(
            0, 0, (1, 1))/3 + sqrt(6)*JzKetCoupled(2, 0, (1, 1))/3
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, -1))) == \
        sqrt(2)*JzKetCoupled(
            1, -1, (1, 1))/2 + sqrt(2)*JzKetCoupled(2, -1, (1, 1))/2
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 1))) == \
        sqrt(3)*JzKetCoupled(0, 0, (1, 1))/3 - sqrt(2)*JzKetCoupled(
            1, 0, (1, 1))/2 + sqrt(6)*JzKetCoupled(2, 0, (1, 1))/6
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 0))) == \
        -sqrt(2)*JzKetCoupled(
            1, -1, (1, 1))/2 + sqrt(2)*JzKetCoupled(2, -1, (1, 1))/2
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, -1))) == \
        JzKetCoupled(2, -2, (1, 1))
    # j1=3/2, j2=1/2
    assert couple(TensorProduct(JzKet(Rational(3, 2), Rational(3, 2)), JzKet(S.Half, S.Half))) == \
        JzKetCoupled(2, 2, (Rational(3, 2), S.Half))
    assert couple(TensorProduct(JzKet(Rational(3, 2), Rational(3, 2)), JzKet(S.Half, Rational(-1, 2)))) == \
        sqrt(3)*JzKetCoupled(
            1, 1, (Rational(3, 2), S.Half))/2 + JzKetCoupled(2, 1, (Rational(3, 2), S.Half))/2
    assert couple(TensorProduct(JzKet(Rational(3, 2), S.Half), JzKet(S.Half, S.Half))) == \
        -JzKetCoupled(1, 1, (S(
            3)/2, S.Half))/2 + sqrt(3)*JzKetCoupled(2, 1, (Rational(3, 2), S.Half))/2
    assert couple(TensorProduct(JzKet(Rational(3, 2), S.Half), JzKet(S.Half, Rational(-1, 2)))) == \
        sqrt(2)*JzKetCoupled(1, 0, (S(
            3)/2, S.Half))/2 + sqrt(2)*JzKetCoupled(2, 0, (Rational(3, 2), S.Half))/2
    assert couple(TensorProduct(JzKet(Rational(3, 2), Rational(-1, 2)), JzKet(S.Half, S.Half))) == \
        -sqrt(2)*JzKetCoupled(1, 0, (S(
            3)/2, S.Half))/2 + sqrt(2)*JzKetCoupled(2, 0, (Rational(3, 2), S.Half))/2
    assert couple(TensorProduct(JzKet(Rational(3, 2), Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))) == \
        JzKetCoupled(1, -1, (S(
            3)/2, S.Half))/2 + sqrt(3)*JzKetCoupled(2, -1, (Rational(3, 2), S.Half))/2
    assert couple(TensorProduct(JzKet(Rational(3, 2), Rational(-3, 2)), JzKet(S.Half, S.Half))) == \
        -sqrt(3)*JzKetCoupled(1, -1, (Rational(3, 2), S.Half))/2 + \
        JzKetCoupled(2, -1, (Rational(3, 2), S.Half))/2
    assert couple(TensorProduct(JzKet(Rational(3, 2), Rational(-3, 2)), JzKet(S.Half, Rational(-1, 2)))) == \
        JzKetCoupled(2, -2, (Rational(3, 2), S.Half))


def test_couple_3_states_numerical():
    # Default coupling
    # j1=1/2,j2=1/2,j3=1/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half))) == \
        JzKetCoupled(Rational(3, 2), S(
            3)/2, (S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2))) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))) == \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half)) )/3 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.One/
             2), ((1, 2, 1), (1, 3, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))) == \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half)) )/2 - \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.One/
             2), ((1, 2, 1), (1, 3, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))) == \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half)) )/2 + \
        sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.One
             /2), ((1, 2, 1), (1, 3, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half))) == \
        -sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half)) )/2 - \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.One/
             2), ((1, 2, 1), (1, 3, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))) == \
        -sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half)) )/2 + \
        sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.One
             /2), ((1, 2, 1), (1, 3, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))) == \
        -sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half)) )/3 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.One
             /2), ((1, 2, 1), (1, 3, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))) == \
        JzKetCoupled(Rational(3, 2), -S(
            3)/2, (S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2))) )
    # j1=S.Half, j2=S.Half, j3=1
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1))) == \
        JzKetCoupled(2, 2, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0))) == \
        sqrt(2)*JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(2)*JzKetCoupled(
            2, 1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1))) == \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 0)) )/3 + \
        sqrt(2)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(6)*JzKetCoupled(
            2, 0, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))) == \
        sqrt(2)*JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 2, 0), (1, 3, 1)) )/2 - \
        JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, 1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))) == \
        -sqrt(6)*JzKetCoupled(0, 0, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 0)) )/6 + \
        sqrt(2)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 2, 0), (1, 3, 1)) )/2 + \
        sqrt(3)*JzKetCoupled(
            2, 0, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )/3
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1))) == \
        sqrt(2)*JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 2, 0), (1, 3, 1)) )/2 + \
        JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, -1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1))) == \
        -sqrt(2)*JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 2, 0), (1, 3, 1)) )/2 - \
        JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, 1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0))) == \
        -sqrt(6)*JzKetCoupled(0, 0, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 0)) )/6 - \
        sqrt(2)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 2, 0), (1, 3, 1)) )/2 + \
        sqrt(3)*JzKetCoupled(
            2, 0, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )/3
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1))) == \
        -sqrt(2)*JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 2, 0), (1, 3, 1)) )/2 + \
        JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, -1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))) == \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 0)) )/3 - \
        sqrt(2)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(6)*JzKetCoupled(
            2, 0, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))) == \
        -sqrt(2)*JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(2)*JzKetCoupled(
            2, -1, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1))) == \
        JzKetCoupled(2, -2, (S.Half, S.Half, 1), ((1, 2, 1), (1, 3, 2)) )
    # j1=S.Half, j2=1, j3=1
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 1))) == \
        JzKetCoupled(
            Rational(5, 2), Rational(5, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(5, 2))) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0))) == \
        sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(S(
            5)/2, Rational(3, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1))) == \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, S.Half)) )/2 + \
        sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, 1, 1), ((1,
             2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1))) == \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        2*sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(S(
            5)/2, Rational(3, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0))) == \
        JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, S.Half)) )/3 - \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(S(
            5)/2, S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1))) == \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, S.Half)) )/3 + \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, S.Half)) )/3 + \
        JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        4*sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1), ((1,
             2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1))) == \
        -2*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, S.Half)) )/6 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, 1, 1), ((1,
             2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0))) == \
        -sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, S.Half)) )/3 - \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, S.Half)) )/3 + \
        2*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1), ((1,
             2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, -1))) == \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, 1, 1), ((1,
             2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1))) == \
        -sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(S(
            5)/2, Rational(3, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0))) == \
        -sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, S.Half)) )/3 - \
        JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, S.Half)) )/3 - \
        2*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(S(
            5)/2, S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1))) == \
        -2*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1), ((1,
             2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1))) == \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, S.Half)) )/3 + \
        JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, S.Half)) )/3 - \
        JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        4*sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(S(
            5)/2, S.Half, (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0))) == \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, S.Half)) )/3 - \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, S.Half)) )/3 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1), ((1,
             2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1))) == \
        -sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 2, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        2*sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, 1, 1), ((1,
             2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1))) == \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, S.Half)) )/2 - \
        sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1), ((1,
             2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0))) == \
        -sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, 1, 1), ((1,
             2, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, -1))) == \
        JzKetCoupled(S(
            5)/2, Rational(-5, 2), (S.Half, 1, 1), ((1, 2, Rational(3, 2)), (1, 3, Rational(5, 2))) )
    #  j1=1, j2=1, j3=1
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, 1))) == \
        JzKetCoupled(3, 3, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, 0))) == \
        sqrt(6)*JzKetCoupled(2, 2, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/3 + \
        sqrt(3)*JzKetCoupled(3, 2, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/3
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, -1))) == \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/5 + \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/3 + \
        sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 1))) == \
        sqrt(2)*JzKetCoupled(2, 2, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 - \
        sqrt(6)*JzKetCoupled(2, 2, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        sqrt(3)*JzKetCoupled(3, 2, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/3
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 0))) == \
        JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 - \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/10 + \
        JzKetCoupled(2, 1, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 + \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        2*sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, -1))) == \
        sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 0)) )/6 + \
        JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/10 + \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/6 + \
        JzKetCoupled(2, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/2 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/10
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 1))) == \
        sqrt(3)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 0), (1, 3, 1)) )/3 - \
        JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/30 + \
        JzKetCoupled(2, 1, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 - \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 0))) == \
        -sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 0)) )/6 + \
        sqrt(3)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 0), (1, 3, 1)) )/3 - \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/15 + \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/3 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/10
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, -1))) == \
        sqrt(3)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 0), (1, 3, 1)) )/3 + \
        JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/30 + \
        JzKetCoupled(2, -1, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 + \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, 1))) == \
        -sqrt(2)*JzKetCoupled(2, 2, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 - \
        sqrt(6)*JzKetCoupled(2, 2, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        sqrt(3)*JzKetCoupled(3, 2, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/3
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, 0))) == \
        -JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 - \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/10 - \
        JzKetCoupled(2, 1, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 + \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        2*sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, -1))) == \
        -sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 0)) )/6 - \
        JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/10 - \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/6 + \
        JzKetCoupled(2, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/2 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/10
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 1))) == \
        -sqrt(3)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 0), (1, 3, 1)) )/3 + \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/15 - \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/3 + \
        2*sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 0))) == \
        -sqrt(3)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 0), (1, 3, 1)) )/3 - \
        2*sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/15 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/5
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, -1))) == \
        -sqrt(3)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 0), (1, 3, 1)) )/3 + \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/15 + \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/3 + \
        2*sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 1))) == \
        sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 0)) )/6 - \
        JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/10 + \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/6 - \
        JzKetCoupled(2, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/2 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/10
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 0))) == \
        -JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 - \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/10 + \
        JzKetCoupled(2, -1, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 - \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        2*sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, -1))) == \
        sqrt(2)*JzKetCoupled(2, -2, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 + \
        sqrt(6)*JzKetCoupled(2, -2, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        sqrt(3)*JzKetCoupled(3, -2, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/3
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 1))) == \
        sqrt(3)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 0), (1, 3, 1)) )/3 + \
        JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/30 - \
        JzKetCoupled(2, 1, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 - \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 0))) == \
        sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 0)) )/6 + \
        sqrt(3)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 0), (1, 3, 1)) )/3 - \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/15 - \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/3 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/10
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, -1))) == \
        sqrt(3)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 0), (1, 3, 1)) )/3 - \
        JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/30 - \
        JzKetCoupled(2, -1, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 + \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 1))) == \
        -sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 0)) )/6 + \
        JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/10 - \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/6 - \
        JzKetCoupled(2, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/2 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/10
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 0))) == \
        JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 1), (1, 3, 1)) )/2 - \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/10 - \
        JzKetCoupled(2, -1, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 - \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        2*sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, -1))) == \
        -sqrt(2)*JzKetCoupled(2, -2, (1, 1, 1), ((1, 2, 1), (1, 3, 2)) )/2 + \
        sqrt(6)*JzKetCoupled(2, -2, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/6 + \
        sqrt(3)*JzKetCoupled(3, -2, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/3
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 1))) == \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 1)) )/5 - \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/3 + \
        sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/15
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 0))) == \
        -sqrt(6)*JzKetCoupled(2, -2, (1, 1, 1), ((1, 2, 2), (1, 3, 2)) )/3 + \
        sqrt(3)*JzKetCoupled(3, -2, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )/3
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, -1))) == \
        JzKetCoupled(3, -3, (1, 1, 1), ((1, 2, 2), (1, 3, 3)) )
    # j1=S.Half, j2=S.Half, j3=Rational(3, 2)
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(3, 2)))) == \
        JzKetCoupled(Rational(5, 2), S(
            5)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(5, 2))) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), S.Half))) == \
        sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/5 + \
        sqrt(15)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S.Half, S(3)
             /2), ((1, 2, 1), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(-1, 2)))) == \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, S.Half)) )/6 + \
        2*sqrt(30)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(-3, 2)))) == \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, S.Half)) )/2 + \
        sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), -S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(3, 2)))) == \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 0), (1, 3, Rational(3, 2))) )/2 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/10 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S.Half, S(3)/
             2), ((1, 2, 1), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), S.Half))) == \
        -sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, S.Half)) )/6 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 0), (1, 3, Rational(3, 2))) )/2 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/30 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(-1, 2)))) == \
        -sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, S.Half)) )/6 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 0), (1, 3, Rational(3, 2))) )/2 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/30 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), -S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(-3, 2)))) == \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 0), (1, 3, Rational(3, 2))) )/2 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/10 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S.Half, S(3)
             /2), ((1, 2, 1), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(3, 2)))) == \
        -sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 0), (1, 3, Rational(3, 2))) )/2 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/10 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S.Half, S(3)/
             2), ((1, 2, 1), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), S.Half))) == \
        -sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 0), (1, 3, Rational(3, 2))) )/2 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/30 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(-1, 2)))) == \
        -sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 0), (1, 3, Rational(3, 2))) )/2 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/30 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), -S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(-3, 2)))) == \
        -sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 0), (1, 3, Rational(3, 2))) )/2 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/10 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S.Half, S(3)
             /2), ((1, 2, 1), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(3, 2)))) == \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, S.Half)) )/2 - \
        sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), S.Half))) == \
        sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, S.Half)) )/6 - \
        2*sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), -S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(-1, 2)))) == \
        -sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(3, 2))) )/5 + \
        sqrt(15)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S.Half, S(
            3)/2), ((1, 2, 1), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(-3, 2)))) == \
        JzKetCoupled(Rational(5, 2), -S(
            5)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 2, 1), (1, 3, Rational(5, 2))) )
    # Couple j1 to j3
    # j1=1/2, j2=1/2, j3=1/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(Rational(3, 2), S(
            3)/2, (S.Half, S.Half, S.Half), ((1, 3, 1), (1, 2, Rational(3, 2))) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) ) == \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 3, 0), (1, 2, S.Half)) )/2 - \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 3, 1), (1, 2, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.One/
             2), ((1, 3, 1), (1, 2, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 3, 1), (1, 2, S.Half)) )/3 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.One/
             2), ((1, 3, 1), (1, 2, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) ) == \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 3, 0), (1, 2, S.Half)) )/2 + \
        sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 3, 1), (1, 2, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.One
             /2), ((1, 3, 1), (1, 2, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) ) == \
        -sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 3, 0), (1, 2, S.Half)) )/2 - \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half), ((1, 3, 1), (1, 2, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.One/
             2), ((1, 3, 1), (1, 2, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) ) == \
        -sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 3, 1), (1, 2, S.Half)) )/3 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.One
             /2), ((1, 3, 1), (1, 2, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 3), (1, 2)) ) == \
        -sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 3, 0), (1, 2, S.Half)) )/2 + \
        sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half), ((1, 3, 1), (1, 2, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.One
             /2), ((1, 3, 1), (1, 2, Rational(3, 2))) )/3
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(Rational(3, 2), -S(
            3)/2, (S.Half, S.Half, S.Half), ((1, 3, 1), (1, 2, Rational(3, 2))) )
    # j1=1/2, j2=1/2, j3=1
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(2, 2, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        sqrt(3)*JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )/3 - \
        sqrt(6)*JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 1)) )/6 + \
        sqrt(2)*JzKetCoupled(
            2, 1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        -sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 0)) )/3 + \
        sqrt(3)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )/3 - \
        sqrt(6)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 1)) )/6 + \
        sqrt(6)*JzKetCoupled(
            2, 0, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        sqrt(3)*JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 1)) )/2 + \
        JzKetCoupled(2, 1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(0, 0, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 0)) )/6 + \
        sqrt(6)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )/6 + \
        sqrt(3)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 1)) )/3 + \
        sqrt(3)*JzKetCoupled(
            2, 0, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )/3
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )/3 + \
        sqrt(3)*JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 1)) )/6 + \
        JzKetCoupled(
            2, -1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        -sqrt(6)*JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )/3 - \
        sqrt(3)*JzKetCoupled(1, 1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 1)) )/6 + \
        JzKetCoupled(2, 1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(0, 0, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 0)) )/6 - \
        sqrt(6)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )/6 - \
        sqrt(3)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 1)) )/3 + \
        sqrt(3)*JzKetCoupled(
            2, 0, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )/3
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        -sqrt(3)*JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 1)) )/2 + \
        JzKetCoupled(
            2, -1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        -sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 0)) )/3 - \
        sqrt(3)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )/3 + \
        sqrt(6)*JzKetCoupled(1, 0, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 1)) )/6 + \
        sqrt(6)*JzKetCoupled(
            2, 0, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        -sqrt(3)*JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 3, S.Half), (1, 2, 1)) )/3 + \
        sqrt(6)*JzKetCoupled(1, -1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 1)) )/6 + \
        sqrt(2)*JzKetCoupled(
            2, -1, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(2, -2, (S.Half, S.Half, 1), ((1, 3, Rational(3, 2)), (1, 2, 2)) )
    # j 1=1/2, j 2=1, j 3=1
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(
            Rational(5, 2), Rational(5, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(5, 2))) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 - \
        2*sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(S(
            5)/2, Rational(3, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 1), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        -2*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, S.Half)) )/6 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 - \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, 1, 1), ((1,
             3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(S(
            5)/2, Rational(3, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, S.Half)) )/3 - \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 + \
        sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(S(
            5)/2, S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, 0), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        -sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, S.Half)) )/3 - \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, S.Half)) )/3 + \
        2*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 - \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1), ((1,
             3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, S.Half)) )/2 + \
        sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, 1, 1), ((1,
             3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, S.Half)) )/3 + \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, S.Half)) )/3 + \
        JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 + \
        4*sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1), ((1,
             3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(1, -1), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, 1, 1), ((1,
             3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        -sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(S(
            5)/2, Rational(3, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, S.Half)) )/3 + \
        JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, S.Half)) )/3 - \
        JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 - \
        4*sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(S(
            5)/2, S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, S.Half)) )/2 - \
        sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1), ((1,
             3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        -sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, S.Half)) )/3 - \
        JzKetCoupled(S.Half, S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, S.Half)) )/3 - \
        2*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 + \
        sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(S(
            5)/2, S.Half, (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, S.Half)) )/3 - \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, S.Half)) )/3 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 - \
        sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1), ((1,
             3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        -sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, 1, 1), ((1,
             3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        -2*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 + \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, 1, 1), ((1,
             3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        -sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 3, S.Half), (1, 2, Rational(3, 2))) )/3 + \
        2*sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, 1, 1), ((1,
             3, Rational(3, 2)), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(S(
            5)/2, Rational(-5, 2), (S.Half, 1, 1), ((1, 3, Rational(3, 2)), (1, 2, Rational(5, 2))) )
    # j1=1, 1, 1
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(3, 3, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        sqrt(2)*JzKetCoupled(2, 2, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 - \
        sqrt(6)*JzKetCoupled(2, 2, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        sqrt(3)*JzKetCoupled(3, 2, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/3
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 1), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        sqrt(3)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 0), (1, 2, 1)) )/3 - \
        JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/30 + \
        JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 - \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(2, 2, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/3 + \
        sqrt(3)*JzKetCoupled(3, 2, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/3
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 - \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/10 + \
        JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 + \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        2*sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, 0), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        -sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 0)) )/6 + \
        sqrt(3)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 0), (1, 2, 1)) )/3 - \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/15 + \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/3 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/10
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/5 + \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/3 + \
        sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 0)) )/6 + \
        JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/10 + \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/6 + \
        JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/2 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/10
    assert couple(TensorProduct(JzKet(1, 1), JzKet(1, -1), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        sqrt(3)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 0), (1, 2, 1)) )/3 + \
        JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/30 + \
        JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 + \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        -sqrt(2)*JzKetCoupled(2, 2, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 - \
        sqrt(6)*JzKetCoupled(2, 2, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        sqrt(3)*JzKetCoupled(3, 2, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/3
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        -sqrt(3)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 0), (1, 2, 1)) )/3 + \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/15 - \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/3 + \
        2*sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 1), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 0)) )/6 - \
        JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/10 + \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/6 - \
        JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/2 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/10
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        -JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 - \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/10 - \
        JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 + \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        2*sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        -sqrt(3)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 0), (1, 2, 1)) )/3 - \
        2*sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/15 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/5
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, 0), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        -JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 - \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/10 + \
        JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 - \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        2*sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        -sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 0)) )/6 - \
        JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/10 - \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/6 + \
        JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/2 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/10
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        -sqrt(3)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 0), (1, 2, 1)) )/3 + \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/15 + \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/3 + \
        2*sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, 0), JzKet(1, -1), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        sqrt(2)*JzKetCoupled(2, -2, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 + \
        sqrt(6)*JzKetCoupled(2, -2, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        sqrt(3)*JzKetCoupled(3, -2, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/3
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        sqrt(3)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 0), (1, 2, 1)) )/3 + \
        JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/30 - \
        JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 - \
        sqrt(3)*JzKetCoupled(2, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        sqrt(15)*JzKetCoupled(3, 1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        -sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 0)) )/6 + \
        JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/10 - \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/6 - \
        JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/2 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/10
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 1), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/5 - \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/3 + \
        sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(0, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 0)) )/6 + \
        sqrt(3)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 0), (1, 2, 1)) )/3 - \
        sqrt(15)*JzKetCoupled(1, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/15 - \
        sqrt(3)*JzKetCoupled(2, 0, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/3 + \
        sqrt(10)*JzKetCoupled(3, 0, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/10
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 - \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/10 - \
        JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 - \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        2*sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, 0), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        -sqrt(6)*JzKetCoupled(2, -2, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/3 + \
        sqrt(3)*JzKetCoupled(3, -2, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/3
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 1)), ((1, 3), (1, 2)) ) == \
        sqrt(3)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 0), (1, 2, 1)) )/3 - \
        JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 1), (1, 2, 1)) )/2 + \
        sqrt(15)*JzKetCoupled(1, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 1)) )/30 - \
        JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 + \
        sqrt(3)*JzKetCoupled(2, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        sqrt(15)*JzKetCoupled(3, -1, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/15
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, 0)), ((1, 3), (1, 2)) ) == \
        -sqrt(2)*JzKetCoupled(2, -2, (1, 1, 1), ((1, 3, 1), (1, 2, 2)) )/2 + \
        sqrt(6)*JzKetCoupled(2, -2, (1, 1, 1), ((1, 3, 2), (1, 2, 2)) )/6 + \
        sqrt(3)*JzKetCoupled(3, -2, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )/3
    assert couple(TensorProduct(JzKet(1, -1), JzKet(1, -1), JzKet(1, -1)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(3, -3, (1, 1, 1), ((1, 3, 2), (1, 2, 3)) )
    # j1=1/2, j2=1/2, j3=3/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(3, 2))), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(Rational(5, 2), S(
            5)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(5, 2))) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), S.Half)), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/2 - \
        sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/10 + \
        sqrt(15)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S.Half, S(3)
             /2), ((1, 3, 2), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(-1, 2))), ((1, 3), (1, 2)) ) == \
        -sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/3 - \
        sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/5 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(-3, 2))), ((1, 3), (1, 2)) ) == \
        -sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, S.Half)) )/2 + \
        JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/2 - \
        sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/10 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), -S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(3, 2))), ((1, 3), (1, 2)) ) == \
        2*sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/5 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S.Half, S(3)/
             2), ((1, 3, 2), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), S.Half)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/6 + \
        3*sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/10 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(-1, 2))), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, S.Half)) )/6 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/3 + \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/5 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), -S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(-3, 2))), ((1, 3), (1, 2)) ) == \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/2 + \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/10 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S.Half, S(3)
             /2), ((1, 3, 2), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(3, 2))), ((1, 3), (1, 2)) ) == \
        -sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/2 - \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/10 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S.Half, S(3)/
             2), ((1, 3, 2), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), S.Half)), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, S.Half)) )/6 - \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/3 - \
        sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/5 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(-1, 2))), ((1, 3), (1, 2)) ) == \
        sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, S.Half)) )/6 - \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/6 - \
        3*sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/10 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), -S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(Rational(3, 2), Rational(-3, 2))), ((1, 3), (1, 2)) ) == \
        -2*sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/5 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S.Half, S(3)
             /2), ((1, 3, 2), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(3, 2))), ((1, 3), (1, 2)) ) == \
        -sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, S.Half)) )/2 - \
        JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/2 + \
        sqrt(15)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/10 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), S.Half)), ((1, 3), (1, 2)) ) == \
        -sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, S.Half)) )/6 - \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/3 + \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/5 + \
        sqrt(30)*JzKetCoupled(Rational(5, 2), -S(
            1)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(-1, 2))), ((1, 3), (1, 2)) ) == \
        -JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 1), (1, 2, Rational(3, 2))) )/2 + \
        sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(3, 2))) )/10 + \
        sqrt(15)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S.Half, S(
            3)/2), ((1, 3, 2), (1, 2, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(Rational(3, 2), Rational(-3, 2))), ((1, 3), (1, 2)) ) == \
        JzKetCoupled(Rational(5, 2), -S(
            5)/2, (S.Half, S.Half, Rational(3, 2)), ((1, 3, 2), (1, 2, Rational(5, 2))) )


def test_couple_4_states_numerical():
    # Default coupling
    # j1=1/2, j2=1/2, j3=1/2, j4=1/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half))) == \
        JzKetCoupled(2, 2, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))) == \
        sqrt(3)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/2 + \
        JzKetCoupled(2, 1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))) == \
        sqrt(6)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/3 - \
        sqrt(3)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        JzKetCoupled(2, 1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))) == \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 0)) )/3 + \
        sqrt(3)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/3 + \
        sqrt(6)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half))) == \
        sqrt(2)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half), (1, 4, 1)) )/2 - \
        sqrt(6)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/6 - \
        sqrt(3)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        JzKetCoupled(2, 1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)),
                        JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))) == \
        JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half),
                ((1, 2, 0), (1, 3, S.Half), (1, 4, 0)))/2 - \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half),
                ((1, 2, 1), (1, 3, S.Half), (1, 4, 0)))/6 + \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half),
                ((1, 2, 0), (1, 3, S.Half), (1, 4, 1)))/2 - \
        sqrt(3)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half),
                ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)))/6 + \
        sqrt(6)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half),
                ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)))/6 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S.Half, S.Half, S.Half),
                ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)))/6
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))) == \
        -JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half), (1, 4, 0)) )/2 - \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 0)) )/6 + \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half), (1, 4, 1)) )/2 + \
        sqrt(3)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/6 - \
        sqrt(6)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))) == \
        sqrt(2)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half), (1, 4, 1)) )/2 + \
        sqrt(6)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/6 + \
        sqrt(3)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        JzKetCoupled(2, -1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half))) == \
        -sqrt(2)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half), (1, 4, 1)) )/2 - \
        sqrt(6)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/6 - \
        sqrt(3)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        JzKetCoupled(2, 1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))) == \
        -JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half), (1, 4, 0)) )/2 - \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 0)) )/6 - \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half), (1, 4, 1)) )/2 - \
        sqrt(3)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/6 + \
        sqrt(6)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))) == \
        JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half), (1, 4, 0)) )/2 - \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 0)) )/6 - \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half), (1, 4, 1)) )/2 + \
        sqrt(3)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/6 - \
        sqrt(6)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))) == \
        -sqrt(2)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (1, 3, S.Half), (1, 4, 1)) )/2 + \
        sqrt(6)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/6 + \
        sqrt(3)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        JzKetCoupled(2, -1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half))) == \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 0)) )/3 - \
        sqrt(3)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/3 - \
        sqrt(6)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)))) == \
        -sqrt(6)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, S.Half), (1, 4, 1)) )/3 + \
        sqrt(3)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/6 + \
        JzKetCoupled(2, -1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half))) == \
        -sqrt(3)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 1)) )/2 + \
        JzKetCoupled(2, -1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)))) == \
        JzKetCoupled(2, -2, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, 2)) )
    # j1=S.Half, S.Half, S.Half, 1
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1))) == \
        JzKetCoupled(Rational(5, 2), Rational(5, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0))) == \
        sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1))) == \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/2 + \
        sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))) == \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))) == \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/3 - \
        JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/3 + \
        2*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 + \
        sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1))) == \
        2*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/6 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 + \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1))) == \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/2 - \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0))) == \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/6 - \
        JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/3 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 - \
        JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 + \
        sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1))) == \
        sqrt(3)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, S.Half)) )/3 - \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/6 + \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 + \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))) == \
        -sqrt(3)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, S.Half)) )/3 - \
        JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/6 + \
        sqrt(6)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 - \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))) == \
        -sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/6 - \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/3 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 + \
        JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 - \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1))) == \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/2 + \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1))) == \
        -sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/2 - \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0))) == \
        -sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/6 - \
        JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/3 - \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 - \
        JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 + \
        sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1))) == \
        -sqrt(3)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, S.Half)) )/3 - \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/6 - \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 + \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))) == \
        sqrt(3)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, S.Half)) )/3 - \
        JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/6 - \
        sqrt(6)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 - \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))) == \
        sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/6 - \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/3 - \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 + \
        JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 - \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1))) == \
        -sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/2 + \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/6 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1))) == \
        2*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 - \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0))) == \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, S.Half)) )/3 - \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/3 - \
        2*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 - \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1))) == \
        -sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, S.Half), (1, 4, Rational(3, 2))) )/3 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1))) == \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, S.Half)) )/2 - \
        sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0))) == \
        -sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1))) == \
        JzKetCoupled(Rational(5, 2), Rational(-5, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (1, 3, Rational(3, 2)), (1, 4, Rational(5, 2))) )
    # Couple j1 to j2, j3 to j4
    # j1=1/2, j2=1/2, j3=1/2, j4=1/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 2), (3, 4), (1, 3)) ) == \
        JzKetCoupled(2, 2, (S(
            1)/2, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(2)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 0), (1, 3, 1)) )/2 + \
        JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, 1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(2)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 0), (1, 3, 1)) )/2 + \
        JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, 1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 0)) )/3 + \
        sqrt(2)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 1)) )/2 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S.Half, S.Half, S.One/
             2), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(2)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 1), (1, 3, 1)) )/2 - \
        JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, 1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 2), (3, 4), (1, 3)) ) == \
        JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 0), (1, 3, 0)) )/2 - \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 0)) )/6 + \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 0), (1, 3, 1)) )/2 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S.Half, S.Half, S.One/
             2), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 2), (3, 4), (1, 3)) ) == \
        -JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 0), (1, 3, 0)) )/2 - \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 0)) )/6 + \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 1), (1, 3, 1)) )/2 - \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 0), (1, 3, 1)) )/2 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S.Half, S.Half, S.One/
             2), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(2)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, -1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(2)*JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 1), (1, 3, 1)) )/2 - \
        JzKetCoupled(1, 1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, 1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 2), (3, 4), (1, 3)) ) == \
        -JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 0), (1, 3, 0)) )/2 - \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 0)) )/6 - \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 0), (1, 3, 1)) )/2 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S.Half, S.Half, S.One/
             2), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 2), (3, 4), (1, 3)) ) == \
        JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 0), (1, 3, 0)) )/2 - \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 0)) )/6 - \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 1), (1, 3, 1)) )/2 - \
        JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 0), (1, 3, 1)) )/2 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S.Half, S.Half, S.One/
             2), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(2)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 0), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, -1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(3)*JzKetCoupled(0, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 0)) )/3 - \
        sqrt(2)*JzKetCoupled(1, 0, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 1)) )/2 + \
        sqrt(6)*JzKetCoupled(2, 0, (S.Half, S.Half, S.Half, S.One/
             2), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/6
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(2)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 0), (1, 3, 1)) )/2 - \
        JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, -1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(2)*JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 0), (1, 3, 1)) )/2 - \
        JzKetCoupled(1, -1, (S.Half, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 1)) )/2 + \
        JzKetCoupled(2, -1, (S.Half, S(
            1)/2, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )/2
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2))), ((1, 2), (3, 4), (1, 3)) ) == \
        JzKetCoupled(2, -2, (S(
            1)/2, S.Half, S.Half, S.Half), ((1, 2, 1), (3, 4, 1), (1, 3, 2)) )
    # j1=S.Half, S.Half, S.Half, 1
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1)), ((1, 2), (3, 4), (1, 3)) ) == \
        JzKetCoupled(Rational(5, 2), Rational(5, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        2*sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1)), ((1, 2), (3, 4), (1, 3)) ) == \
        2*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/6 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/3 + \
        JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/3 - \
        JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        4*sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/2 + \
        sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/2 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/10 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, S.Half), (1, 3, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/6 - \
        JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/3 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/3 + \
        JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(3)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, S.Half), (1, 3, S.Half)) )/3 + \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/3 - \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/6 + \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/6 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/30 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(3)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, S.Half), (1, 3, S.Half)) )/3 + \
        JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/3 - \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/6 + \
        sqrt(6)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/6 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/30 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, S.Half), (1, 3, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/6 - \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/3 + \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/3 - \
        JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/2 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/10 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 1)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/2 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/10 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, 0)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(6)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, S.Half), (1, 3, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/6 - \
        JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/3 - \
        sqrt(3)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/3 + \
        JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        sqrt(5)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, S.Half), JzKet(1, -1)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(3)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, S.Half), (1, 3, S.Half)) )/3 + \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/3 - \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/6 - \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/6 + \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/30 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(3)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, S.Half), (1, 3, S.Half)) )/3 + \
        JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/3 - \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/6 - \
        sqrt(6)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/6 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/30 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(6)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, S.Half), (1, 3, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/6 - \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/3 - \
        sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/3 - \
        JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 + \
        sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 0), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/2 + \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/10 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 1)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(2)*JzKetCoupled(S.Half, S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/2 - \
        sqrt(10)*JzKetCoupled(Rational(3, 2), S.Half, (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/5 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), S.Half, (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, 0)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/3 + \
        JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/3 + \
        JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        4*sqrt(5)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, S.Half), JzKet(1, -1)), ((1, 2), (3, 4), (1, 3)) ) == \
        sqrt(6)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        sqrt(30)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(5)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 1)), ((1, 2), (3, 4), (1, 3)) ) == \
        2*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, S.Half)) )/3 + \
        sqrt(2)*JzKetCoupled(S.Half, Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, S.Half)) )/6 - \
        sqrt(2)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        2*sqrt(10)*JzKetCoupled(Rational(3, 2), Rational(-1, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-1, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/10
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, 0)), ((1, 2), (3, 4), (1, 3)) ) == \
        -sqrt(3)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, S.Half), (1, 3, Rational(3, 2))) )/3 - \
        2*sqrt(15)*JzKetCoupled(Rational(3, 2), Rational(-3, 2), (S.Half, S.Half, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(3, 2))) )/15 + \
        sqrt(10)*JzKetCoupled(Rational(5, 2), Rational(-3, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )/5
    assert couple(TensorProduct(JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(S.Half, Rational(-1, 2)), JzKet(1, -1)), ((1, 2), (3, 4), (1, 3)) ) == \
        JzKetCoupled(Rational(5, 2), Rational(-5, 2), (S.Half, S(
            1)/2, S.Half, 1), ((1, 2, 1), (3, 4, Rational(3, 2)), (1, 3, Rational(5, 2))) )


def test_couple_symbolic():
    assert couple(TensorProduct(JzKet(j1, m1), JzKet(j2, m2))) == \
        Sum(CG(j1, m1, j2, m2, j, m1 + m2) * JzKetCoupled(j, m1 + m2, (
            j1, j2)), (j, m1 + m2, j1 + j2))
    assert couple(TensorProduct(JzKet(j1, m1), JzKet(j2, m2), JzKet(j3, m3))) == \
        Sum(CG(j1, m1, j2, m2, j12, m1 + m2) * CG(j12, m1 + m2, j3, m3, j, m1 + m2 + m3) *
            JzKetCoupled(j, m1 + m2 + m3, (j1, j2, j3), ((1, 2, j12), (1, 3, j)) ),
            (j12, m1 + m2, j1 + j2), (j, m1 + m2 + m3, j12 + j3))
    assert couple(TensorProduct(JzKet(j1, m1), JzKet(j2, m2), JzKet(j3, m3)), ((1, 3), (1, 2)) ) == \
        Sum(CG(j1, m1, j3, m3, j13, m1 + m3) * CG(j13, m1 + m3, j2, m2, j, m1 + m2 + m3) *
            JzKetCoupled(j, m1 + m2 + m3, (j1, j2, j3), ((1, 3, j13), (1, 2, j)) ),
            (j13, m1 + m3, j1 + j3), (j, m1 + m2 + m3, j13 + j2))
    assert couple(TensorProduct(JzKet(j1, m1), JzKet(j2, m2), JzKet(j3, m3), JzKet(j4, m4))) == \
        Sum(CG(j1, m1, j2, m2, j12, m1 + m2) * CG(j12, m1 + m2, j3, m3, j123, m1 + m2 + m3) * CG(j123, m1 + m2 + m3, j4, m4, j, m1 + m2 + m3 + m4) *
            JzKetCoupled(j, m1 + m2 + m3 + m4, (
                j1, j2, j3, j4), ((1, 2, j12), (1, 3, j123), (1, 4, j)) ),
            (j12, m1 + m2, j1 + j2), (j123, m1 + m2 + m3, j12 + j3), (j, m1 + m2 + m3 + m4, j123 + j4))
    assert couple(TensorProduct(JzKet(j1, m1), JzKet(j2, m2), JzKet(j3, m3), JzKet(j4, m4)), ((1, 2), (3, 4), (1, 3)) ) == \
        Sum(CG(j1, m1, j2, m2, j12, m1 + m2) * CG(j3, m3, j4, m4, j34, m3 + m4) * CG(j12, m1 + m2, j34, m3 + m4, j, m1 + m2 + m3 + m4) *
            JzKetCoupled(j, m1 + m2 + m3 + m4, (
                j1, j2, j3, j4), ((1, 2, j12), (3, 4, j34), (1, 3, j)) ),
            (j12, m1 + m2, j1 + j2), (j34, m3 + m4, j3 + j4), (j, m1 + m2 + m3 + m4, j12 + j34))
    assert couple(TensorProduct(JzKet(j1, m1), JzKet(j2, m2), JzKet(j3, m3), JzKet(j4, m4)), ((1, 3), (1, 4), (1, 2)) ) == \
        Sum(CG(j1, m1, j3, m3, j13, m1 + m3) * CG(j13, m1 + m3, j4, m4, j134, m1 + m3 + m4) * CG(j134, m1 + m3 + m4, j2, m2, j, m1 + m2 + m3 + m4) *
            JzKetCoupled(j, m1 + m2 + m3 + m4, (
                j1, j2, j3, j4), ((1, 3, j13), (1, 4, j134), (1, 2, j)) ),
            (j13, m1 + m3, j1 + j3), (j134, m1 + m3 + m4, j13 + j4), (j, m1 + m2 + m3 + m4, j134 + j2))


def test_innerproduct():
    assert InnerProduct(JzBra(1, 1), JzKet(1, 1)).doit() == 1
    assert InnerProduct(
        JzBra(S.Half, S.Half), JzKet(S.Half, Rational(-1, 2))).doit() == 0
    assert InnerProduct(JzBra(j, m), JzKet(j, m)).doit() == 1
    assert InnerProduct(JzBra(1, 0), JyKet(1, 1)).doit() == I/sqrt(2)
    assert InnerProduct(
        JxBra(S.Half, S.Half), JzKet(S.Half, S.Half)).doit() == -sqrt(2)/2
    assert InnerProduct(JyBra(1, 1), JzKet(1, 1)).doit() == S.Half
    assert InnerProduct(JxBra(1, -1), JyKet(1, 1)).doit() == 0


def test_rotation_small_d():
    # Symbolic tests
    # j = 1/2
    assert Rotation.d(S.Half, S.Half, S.Half, beta).doit() == cos(beta/2)
    assert Rotation.d(S.Half, S.Half, Rational(-1, 2), beta).doit() == -sin(beta/2)
    assert Rotation.d(S.Half, Rational(-1, 2), S.Half, beta).doit() == sin(beta/2)
    assert Rotation.d(S.Half, Rational(-1, 2), Rational(-1, 2), beta).doit() == cos(beta/2)
    # j = 1
    assert Rotation.d(1, 1, 1, beta).doit() == (1 + cos(beta))/2
    assert Rotation.d(1, 1, 0, beta).doit() == -sin(beta)/sqrt(2)
    assert Rotation.d(1, 1, -1, beta).doit() == (1 - cos(beta))/2
    assert Rotation.d(1, 0, 1, beta).doit() == sin(beta)/sqrt(2)
    assert Rotation.d(1, 0, 0, beta).doit() == cos(beta)
    assert Rotation.d(1, 0, -1, beta).doit() == -sin(beta)/sqrt(2)
    assert Rotation.d(1, -1, 1, beta).doit() == (1 - cos(beta))/2
    assert Rotation.d(1, -1, 0, beta).doit() == sin(beta)/sqrt(2)
    assert Rotation.d(1, -1, -1, beta).doit() == (1 + cos(beta))/2
    # j = 3/2
    assert Rotation.d(S(
        3)/2, Rational(3, 2), Rational(3, 2), beta).doit() == (3*cos(beta/2) + cos(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), S(
        3)/2, S.Half, beta).doit() == -sqrt(3)*(sin(beta/2) + sin(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), S(
        3)/2, Rational(-1, 2), beta).doit() == sqrt(3)*(cos(beta/2) - cos(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), S(
        3)/2, Rational(-3, 2), beta).doit() == (-3*sin(beta/2) + sin(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), S(
        1)/2, Rational(3, 2), beta).doit() == sqrt(3)*(sin(beta/2) + sin(beta*Rational(3, 2)))/4
    assert Rotation.d(S(
        3)/2, S.Half, S.Half, beta).doit() == (cos(beta/2) + 3*cos(beta*Rational(3, 2)))/4
    assert Rotation.d(S(
        3)/2, S.Half, Rational(-1, 2), beta).doit() == (sin(beta/2) - 3*sin(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), S(
        1)/2, Rational(-3, 2), beta).doit() == sqrt(3)*(cos(beta/2) - cos(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), -S(
        1)/2, Rational(3, 2), beta).doit() == sqrt(3)*(cos(beta/2) - cos(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), -S(
        1)/2, S.Half, beta).doit() == (-sin(beta/2) + 3*sin(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), -S(
        1)/2, Rational(-1, 2), beta).doit() == (cos(beta/2) + 3*cos(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), -S(
        1)/2, Rational(-3, 2), beta).doit() == -sqrt(3)*(sin(beta/2) + sin(beta*Rational(3, 2)))/4
    assert Rotation.d(S(
        3)/2, Rational(-3, 2), Rational(3, 2), beta).doit() == (3*sin(beta/2) - sin(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), -S(
        3)/2, S.Half, beta).doit() == sqrt(3)*(cos(beta/2) - cos(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), -S(
        3)/2, Rational(-1, 2), beta).doit() == sqrt(3)*(sin(beta/2) + sin(beta*Rational(3, 2)))/4
    assert Rotation.d(Rational(3, 2), -S(
        3)/2, Rational(-3, 2), beta).doit() == (3*cos(beta/2) + cos(beta*Rational(3, 2)))/4
    # j = 2
    assert Rotation.d(2, 2, 2, beta).doit() == (3 + 4*cos(beta) + cos(2*beta))/8
    assert Rotation.d(2, 2, 1, beta).doit() == -((cos(beta) + 1)*sin(beta))/2
    assert Rotation.d(2, 2, 0, beta).doit() == sqrt(6)*sin(beta)**2/4
    assert Rotation.d(2, 2, -1, beta).doit() == (cos(beta) - 1)*sin(beta)/2
    assert Rotation.d(2, 2, -2, beta).doit() == (3 - 4*cos(beta) + cos(2*beta))/8
    assert Rotation.d(2, 1, 2, beta).doit() == (cos(beta) + 1)*sin(beta)/2
    assert Rotation.d(2, 1, 1, beta).doit() == (cos(beta) + cos(2*beta))/2
    assert Rotation.d(2, 1, 0, beta).doit() == -sqrt(6)*sin(2*beta)/4
    assert Rotation.d(2, 1, -1, beta).doit() == (cos(beta) - cos(2*beta))/2
    assert Rotation.d(2, 1, -2, beta).doit() == (cos(beta) - 1)*sin(beta)/2
    assert Rotation.d(2, 0, 2, beta).doit() == sqrt(6)*sin(beta)**2/4
    assert Rotation.d(2, 0, 1, beta).doit() == sqrt(6)*sin(2*beta)/4
    assert Rotation.d(2, 0, 0, beta).doit() == (1 + 3*cos(2*beta))/4
    assert Rotation.d(2, 0, -1, beta).doit() == -sqrt(6)*sin(2*beta)/4
    assert Rotation.d(2, 0, -2, beta).doit() == sqrt(6)*sin(beta)**2/4
    assert Rotation.d(2, -1, 2, beta).doit() == (2*sin(beta) - sin(2*beta))/4
    assert Rotation.d(2, -1, 1, beta).doit() == (cos(beta) - cos(2*beta))/2
    assert Rotation.d(2, -1, 0, beta).doit() == sqrt(6)*sin(2*beta)/4
    assert Rotation.d(2, -1, -1, beta).doit() == (cos(beta) + cos(2*beta))/2
    assert Rotation.d(2, -1, -2, beta).doit() == -((cos(beta) + 1)*sin(beta))/2
    assert Rotation.d(2, -2, 2, beta).doit() == (3 - 4*cos(beta) + cos(2*beta))/8
    assert Rotation.d(2, -2, 1, beta).doit() == (2*sin(beta) - sin(2*beta))/4
    assert Rotation.d(2, -2, 0, beta).doit() == sqrt(6)*sin(beta)**2/4
    assert Rotation.d(2, -2, -1, beta).doit() == (cos(beta) + 1)*sin(beta)/2
    assert Rotation.d(2, -2, -2, beta).doit() == (3 + 4*cos(beta) + cos(2*beta))/8
    # Numerical tests
    # j = 1/2
    assert Rotation.d(S.Half, S.Half, S.Half, pi/2).doit() == sqrt(2)/2
    assert Rotation.d(S.Half, S.Half, Rational(-1, 2), pi/2).doit() == -sqrt(2)/2
    assert Rotation.d(S.Half, Rational(-1, 2), S.Half, pi/2).doit() == sqrt(2)/2
    assert Rotation.d(S.Half, Rational(-1, 2), Rational(-1, 2), pi/2).doit() == sqrt(2)/2
    # j = 1
    assert Rotation.d(1, 1, 1, pi/2).doit() == S.Half
    assert Rotation.d(1, 1, 0, pi/2).doit() == -sqrt(2)/2
    assert Rotation.d(1, 1, -1, pi/2).doit() == S.Half
    assert Rotation.d(1, 0, 1, pi/2).doit() == sqrt(2)/2
    assert Rotation.d(1, 0, 0, pi/2).doit() == 0
    assert Rotation.d(1, 0, -1, pi/2).doit() == -sqrt(2)/2
    assert Rotation.d(1, -1, 1, pi/2).doit() == S.Half
    assert Rotation.d(1, -1, 0, pi/2).doit() == sqrt(2)/2
    assert Rotation.d(1, -1, -1, pi/2).doit() == S.Half
    # j = 3/2
    assert Rotation.d(Rational(3, 2), Rational(3, 2), Rational(3, 2), pi/2).doit() == sqrt(2)/4
    assert Rotation.d(Rational(3, 2), Rational(3, 2), S.Half, pi/2).doit() == -sqrt(6)/4
    assert Rotation.d(Rational(3, 2), Rational(3, 2), Rational(-1, 2), pi/2).doit() == sqrt(6)/4
    assert Rotation.d(Rational(3, 2), Rational(3, 2), Rational(-3, 2), pi/2).doit() == -sqrt(2)/4
    assert Rotation.d(Rational(3, 2), S.Half, Rational(3, 2), pi/2).doit() == sqrt(6)/4
    assert Rotation.d(Rational(3, 2), S.Half, S.Half, pi/2).doit() == -sqrt(2)/4
    assert Rotation.d(Rational(3, 2), S.Half, Rational(-1, 2), pi/2).doit() == -sqrt(2)/4
    assert Rotation.d(Rational(3, 2), S.Half, Rational(-3, 2), pi/2).doit() == sqrt(6)/4
    assert Rotation.d(Rational(3, 2), Rational(-1, 2), Rational(3, 2), pi/2).doit() == sqrt(6)/4
    assert Rotation.d(Rational(3, 2), Rational(-1, 2), S.Half, pi/2).doit() == sqrt(2)/4
    assert Rotation.d(Rational(3, 2), Rational(-1, 2), Rational(-1, 2), pi/2).doit() == -sqrt(2)/4
    assert Rotation.d(Rational(3, 2), Rational(-1, 2), Rational(-3, 2), pi/2).doit() == -sqrt(6)/4
    assert Rotation.d(Rational(3, 2), Rational(-3, 2), Rational(3, 2), pi/2).doit() == sqrt(2)/4
    assert Rotation.d(Rational(3, 2), Rational(-3, 2), S.Half, pi/2).doit() == sqrt(6)/4
    assert Rotation.d(Rational(3, 2), Rational(-3, 2), Rational(-1, 2), pi/2).doit() == sqrt(6)/4
    assert Rotation.d(Rational(3, 2), Rational(-3, 2), Rational(-3, 2), pi/2).doit() == sqrt(2)/4
    # j = 2
    assert Rotation.d(2, 2, 2, pi/2).doit() == Rational(1, 4)
    assert Rotation.d(2, 2, 1, pi/2).doit() == Rational(-1, 2)
    assert Rotation.d(2, 2, 0, pi/2).doit() == sqrt(6)/4
    assert Rotation.d(2, 2, -1, pi/2).doit() == Rational(-1, 2)
    assert Rotation.d(2, 2, -2, pi/2).doit() == Rational(1, 4)
    assert Rotation.d(2, 1, 2, pi/2).doit() == S.Half
    assert Rotation.d(2, 1, 1, pi/2).doit() == Rational(-1, 2)
    assert Rotation.d(2, 1, 0, pi/2).doit() == 0
    assert Rotation.d(2, 1, -1, pi/2).doit() == S.Half
    assert Rotation.d(2, 1, -2, pi/2).doit() == Rational(-1, 2)
    assert Rotation.d(2, 0, 2, pi/2).doit() == sqrt(6)/4
    assert Rotation.d(2, 0, 1, pi/2).doit() == 0
    assert Rotation.d(2, 0, 0, pi/2).doit() == Rational(-1, 2)
    assert Rotation.d(2, 0, -1, pi/2).doit() == 0
    assert Rotation.d(2, 0, -2, pi/2).doit() == sqrt(6)/4
    assert Rotation.d(2, -1, 2, pi/2).doit() == S.Half
    assert Rotation.d(2, -1, 1, pi/2).doit() == S.Half
    assert Rotation.d(2, -1, 0, pi/2).doit() == 0
    assert Rotation.d(2, -1, -1, pi/2).doit() == Rational(-1, 2)
    assert Rotation.d(2, -1, -2, pi/2).doit() == Rational(-1, 2)
    assert Rotation.d(2, -2, 2, pi/2).doit() == Rational(1, 4)
    assert Rotation.d(2, -2, 1, pi/2).doit() == S.Half
    assert Rotation.d(2, -2, 0, pi/2).doit() == sqrt(6)/4
    assert Rotation.d(2, -2, -1, pi/2).doit() == S.Half
    assert Rotation.d(2, -2, -2, pi/2).doit() == Rational(1, 4)


def test_rotation_d():
    # Symbolic tests
    # j = 1/2
    assert Rotation.D(S.Half, S.Half, S.Half, alpha, beta, gamma).doit() == \
        cos(beta/2)*exp(-I*alpha/2)*exp(-I*gamma/2)
    assert Rotation.D(S.Half, S.Half, Rational(-1, 2), alpha, beta, gamma).doit() == \
        -sin(beta/2)*exp(-I*alpha/2)*exp(I*gamma/2)
    assert Rotation.D(S.Half, Rational(-1, 2), S.Half, alpha, beta, gamma).doit() == \
        sin(beta/2)*exp(I*alpha/2)*exp(-I*gamma/2)
    assert Rotation.D(S.Half, Rational(-1, 2), Rational(-1, 2), alpha, beta, gamma).doit() == \
        cos(beta/2)*exp(I*alpha/2)*exp(I*gamma/2)
    # j = 1
    assert Rotation.D(1, 1, 1, alpha, beta, gamma).doit() == \
        (1 + cos(beta))/2*exp(-I*alpha)*exp(-I*gamma)
    assert Rotation.D(1, 1, 0, alpha, beta, gamma).doit() == -sin(
        beta)/sqrt(2)*exp(-I*alpha)
    assert Rotation.D(1, 1, -1, alpha, beta, gamma).doit() == \
        (1 - cos(beta))/2*exp(-I*alpha)*exp(I*gamma)
    assert Rotation.D(1, 0, 1, alpha, beta, gamma).doit() == \
        sin(beta)/sqrt(2)*exp(-I*gamma)
    assert Rotation.D(1, 0, 0, alpha, beta, gamma).doit() == cos(beta)
    assert Rotation.D(1, 0, -1, alpha, beta, gamma).doit() == \
        -sin(beta)/sqrt(2)*exp(I*gamma)
    assert Rotation.D(1, -1, 1, alpha, beta, gamma).doit() == \
        (1 - cos(beta))/2*exp(I*alpha)*exp(-I*gamma)
    assert Rotation.D(1, -1, 0, alpha, beta, gamma).doit() == \
        sin(beta)/sqrt(2)*exp(I*alpha)
    assert Rotation.D(1, -1, -1, alpha, beta, gamma).doit() == \
        (1 + cos(beta))/2*exp(I*alpha)*exp(I*gamma)
    # j = 3/2
    assert Rotation.D(Rational(3, 2), Rational(3, 2), Rational(3, 2), alpha, beta, gamma).doit() == \
        (3*cos(beta/2) + cos(beta*Rational(3, 2)))/4*exp(I*alpha*Rational(-3, 2))*exp(I*gamma*Rational(-3, 2))
    assert Rotation.D(Rational(3, 2), Rational(3, 2), S.Half, alpha, beta, gamma).doit() == \
        -sqrt(3)*(sin(beta/2) + sin(beta*Rational(3, 2)))/4*exp(I*alpha*Rational(-3, 2))*exp(-I*gamma/2)
    assert Rotation.D(Rational(3, 2), Rational(3, 2), Rational(-1, 2), alpha, beta, gamma).doit() == \
        sqrt(3)*(cos(beta/2) - cos(beta*Rational(3, 2)))/4*exp(I*alpha*Rational(-3, 2))*exp(I*gamma/2)
    assert Rotation.D(Rational(3, 2), Rational(3, 2), Rational(-3, 2), alpha, beta, gamma).doit() == \
        (-3*sin(beta/2) + sin(beta*Rational(3, 2)))/4*exp(I*alpha*Rational(-3, 2))*exp(I*gamma*Rational(3, 2))
    assert Rotation.D(Rational(3, 2), S.Half, Rational(3, 2), alpha, beta, gamma).doit() == \
        sqrt(3)*(sin(beta/2) + sin(beta*Rational(3, 2)))/4*exp(-I*alpha/2)*exp(I*gamma*Rational(-3, 2))
    assert Rotation.D(Rational(3, 2), S.Half, S.Half, alpha, beta, gamma).doit() == \
        (cos(beta/2) + 3*cos(beta*Rational(3, 2)))/4*exp(-I*alpha/2)*exp(-I*gamma/2)
    assert Rotation.D(Rational(3, 2), S.Half, Rational(-1, 2), alpha, beta, gamma).doit() == \
        (sin(beta/2) - 3*sin(beta*Rational(3, 2)))/4*exp(-I*alpha/2)*exp(I*gamma/2)
    assert Rotation.D(Rational(3, 2), S.Half, Rational(-3, 2), alpha, beta, gamma).doit() == \
        sqrt(3)*(cos(beta/2) - cos(beta*Rational(3, 2)))/4*exp(-I*alpha/2)*exp(I*gamma*Rational(3, 2))
    assert Rotation.D(Rational(3, 2), Rational(-1, 2), Rational(3, 2), alpha, beta, gamma).doit() == \
        sqrt(3)*(cos(beta/2) - cos(beta*Rational(3, 2)))/4*exp(I*alpha/2)*exp(I*gamma*Rational(-3, 2))
    assert Rotation.D(Rational(3, 2), Rational(-1, 2), S.Half, alpha, beta, gamma).doit() == \
        (-sin(beta/2) + 3*sin(beta*Rational(3, 2)))/4*exp(I*alpha/2)*exp(-I*gamma/2)
    assert Rotation.D(Rational(3, 2), Rational(-1, 2), Rational(-1, 2), alpha, beta, gamma).doit() == \
        (cos(beta/2) + 3*cos(beta*Rational(3, 2)))/4*exp(I*alpha/2)*exp(I*gamma/2)
    assert Rotation.D(Rational(3, 2), Rational(-1, 2), Rational(-3, 2), alpha, beta, gamma).doit() == \
        -sqrt(3)*(sin(beta/2) + sin(beta*Rational(3, 2)))/4*exp(I*alpha/2)*exp(I*gamma*Rational(3, 2))
    assert Rotation.D(Rational(3, 2), Rational(-3, 2), Rational(3, 2), alpha, beta, gamma).doit() == \
        (3*sin(beta/2) - sin(beta*Rational(3, 2)))/4*exp(I*alpha*Rational(3, 2))*exp(I*gamma*Rational(-3, 2))
    assert Rotation.D(Rational(3, 2), Rational(-3, 2), S.Half, alpha, beta, gamma).doit() == \
        sqrt(3)*(cos(beta/2) - cos(beta*Rational(3, 2)))/4*exp(I*alpha*Rational(3, 2))*exp(-I*gamma/2)
    assert Rotation.D(Rational(3, 2), Rational(-3, 2), Rational(-1, 2), alpha, beta, gamma).doit() == \
        sqrt(3)*(sin(beta/2) + sin(beta*Rational(3, 2)))/4*exp(I*alpha*Rational(3, 2))*exp(I*gamma/2)
    assert Rotation.D(Rational(3, 2), Rational(-3, 2), Rational(-3, 2), alpha, beta, gamma).doit() == \
        (3*cos(beta/2) + cos(beta*Rational(3, 2)))/4*exp(I*alpha*Rational(3, 2))*exp(I*gamma*Rational(3, 2))
    # j = 2
    assert Rotation.D(2, 2, 2, alpha, beta, gamma).doit() == \
        (3 + 4*cos(beta) + cos(2*beta))/8*exp(-2*I*alpha)*exp(-2*I*gamma)
    assert Rotation.D(2, 2, 1, alpha, beta, gamma).doit() == \
        -((cos(beta) + 1)*exp(-2*I*alpha)*exp(-I*gamma)*sin(beta))/2
    assert Rotation.D(2, 2, 0, alpha, beta, gamma).doit() == \
        sqrt(6)*sin(beta)**2/4*exp(-2*I*alpha)
    assert Rotation.D(2, 2, -1, alpha, beta, gamma).doit() == \
        (cos(beta) - 1)*sin(beta)/2*exp(-2*I*alpha)*exp(I*gamma)
    assert Rotation.D(2, 2, -2, alpha, beta, gamma).doit() == \
        (3 - 4*cos(beta) + cos(2*beta))/8*exp(-2*I*alpha)*exp(2*I*gamma)
    assert Rotation.D(2, 1, 2, alpha, beta, gamma).doit() == \
        (cos(beta) + 1)*sin(beta)/2*exp(-I*alpha)*exp(-2*I*gamma)
    assert Rotation.D(2, 1, 1, alpha, beta, gamma).doit() == \
        (cos(beta) + cos(2*beta))/2*exp(-I*alpha)*exp(-I*gamma)
    assert Rotation.D(2, 1, 0, alpha, beta, gamma).doit() == -sqrt(6)* \
        sin(2*beta)/4*exp(-I*alpha)
    assert Rotation.D(2, 1, -1, alpha, beta, gamma).doit() == \
        (cos(beta) - cos(2*beta))/2*exp(-I*alpha)*exp(I*gamma)
    assert Rotation.D(2, 1, -2, alpha, beta, gamma).doit() == \
        (cos(beta) - 1)*sin(beta)/2*exp(-I*alpha)*exp(2*I*gamma)
    assert Rotation.D(2, 0, 2, alpha, beta, gamma).doit() == \
        sqrt(6)*sin(beta)**2/4*exp(-2*I*gamma)
    assert Rotation.D(2, 0, 1, alpha, beta, gamma).doit() == sqrt(6)* \
        sin(2*beta)/4*exp(-I*gamma)
    assert Rotation.D(
        2, 0, 0, alpha, beta, gamma).doit() == (1 + 3*cos(2*beta))/4
    assert Rotation.D(2, 0, -1, alpha, beta, gamma).doit() == -sqrt(6)* \
        sin(2*beta)/4*exp(I*gamma)
    assert Rotation.D(2, 0, -2, alpha, beta, gamma).doit() == \
        sqrt(6)*sin(beta)**2/4*exp(2*I*gamma)
    assert Rotation.D(2, -1, 2, alpha, beta, gamma).doit() == \
        (2*sin(beta) - sin(2*beta))/4*exp(I*alpha)*exp(-2*I*gamma)
    assert Rotation.D(2, -1, 1, alpha, beta, gamma).doit() == \
        (cos(beta) - cos(2*beta))/2*exp(I*alpha)*exp(-I*gamma)
    assert Rotation.D(2, -1, 0, alpha, beta, gamma).doit() == sqrt(6)* \
        sin(2*beta)/4*exp(I*alpha)
    assert Rotation.D(2, -1, -1, alpha, beta, gamma).doit() == \
        (cos(beta) + cos(2*beta))/2*exp(I*alpha)*exp(I*gamma)
    assert Rotation.D(2, -1, -2, alpha, beta, gamma).doit() == \
        -((cos(beta) + 1)*sin(beta))/2*exp(I*alpha)*exp(2*I*gamma)
    assert Rotation.D(2, -2, 2, alpha, beta, gamma).doit() == \
        (3 - 4*cos(beta) + cos(2*beta))/8*exp(2*I*alpha)*exp(-2*I*gamma)
    assert Rotation.D(2, -2, 1, alpha, beta, gamma).doit() == \
        (2*sin(beta) - sin(2*beta))/4*exp(2*I*alpha)*exp(-I*gamma)
    assert Rotation.D(2, -2, 0, alpha, beta, gamma).doit() == \
        sqrt(6)*sin(beta)**2/4*exp(2*I*alpha)
    assert Rotation.D(2, -2, -1, alpha, beta, gamma).doit() == \
        (cos(beta) + 1)*sin(beta)/2*exp(2*I*alpha)*exp(I*gamma)
    assert Rotation.D(2, -2, -2, alpha, beta, gamma).doit() == \
        (3 + 4*cos(beta) + cos(2*beta))/8*exp(2*I*alpha)*exp(2*I*gamma)
    # Numerical tests
    # j = 1/2
    assert Rotation.D(
        S.Half, S.Half, S.Half, pi/2, pi/2, pi/2).doit() == -I*sqrt(2)/2
    assert Rotation.D(
        S.Half, S.Half, Rational(-1, 2), pi/2, pi/2, pi/2).doit() == -sqrt(2)/2
    assert Rotation.D(
        S.Half, Rational(-1, 2), S.Half, pi/2, pi/2, pi/2).doit() == sqrt(2)/2
    assert Rotation.D(
        S.Half, Rational(-1, 2), Rational(-1, 2), pi/2, pi/2, pi/2).doit() == I*sqrt(2)/2
    # j = 1
    assert Rotation.D(1, 1, 1, pi/2, pi/2, pi/2).doit() == Rational(-1, 2)
    assert Rotation.D(1, 1, 0, pi/2, pi/2, pi/2).doit() == I*sqrt(2)/2
    assert Rotation.D(1, 1, -1, pi/2, pi/2, pi/2).doit() == S.Half
    assert Rotation.D(1, 0, 1, pi/2, pi/2, pi/2).doit() == -I*sqrt(2)/2
    assert Rotation.D(1, 0, 0, pi/2, pi/2, pi/2).doit() == 0
    assert Rotation.D(1, 0, -1, pi/2, pi/2, pi/2).doit() == -I*sqrt(2)/2
    assert Rotation.D(1, -1, 1, pi/2, pi/2, pi/2).doit() == S.Half
    assert Rotation.D(1, -1, 0, pi/2, pi/2, pi/2).doit() == I*sqrt(2)/2
    assert Rotation.D(1, -1, -1, pi/2, pi/2, pi/2).doit() == Rational(-1, 2)
    # j = 3/2
    assert Rotation.D(
        Rational(3, 2), Rational(3, 2), Rational(3, 2), pi/2, pi/2, pi/2).doit() == I*sqrt(2)/4
    assert Rotation.D(
        Rational(3, 2), Rational(3, 2), S.Half, pi/2, pi/2, pi/2).doit() == sqrt(6)/4
    assert Rotation.D(
        Rational(3, 2), Rational(3, 2), Rational(-1, 2), pi/2, pi/2, pi/2).doit() == -I*sqrt(6)/4
    assert Rotation.D(
        Rational(3, 2), Rational(3, 2), Rational(-3, 2), pi/2, pi/2, pi/2).doit() == -sqrt(2)/4
    assert Rotation.D(
        Rational(3, 2), S.Half, Rational(3, 2), pi/2, pi/2, pi/2).doit() == -sqrt(6)/4
    assert Rotation.D(
        Rational(3, 2), S.Half, S.Half, pi/2, pi/2, pi/2).doit() == I*sqrt(2)/4
    assert Rotation.D(
        Rational(3, 2), S.Half, Rational(-1, 2), pi/2, pi/2, pi/2).doit() == -sqrt(2)/4
    assert Rotation.D(
        Rational(3, 2), S.Half, Rational(-3, 2), pi/2, pi/2, pi/2).doit() == I*sqrt(6)/4
    assert Rotation.D(
        Rational(3, 2), Rational(-1, 2), Rational(3, 2), pi/2, pi/2, pi/2).doit() == -I*sqrt(6)/4
    assert Rotation.D(
        Rational(3, 2), Rational(-1, 2), S.Half, pi/2, pi/2, pi/2).doit() == sqrt(2)/4
    assert Rotation.D(
        Rational(3, 2), Rational(-1, 2), Rational(-1, 2), pi/2, pi/2, pi/2).doit() == -I*sqrt(2)/4
    assert Rotation.D(
        Rational(3, 2), Rational(-1, 2), Rational(-3, 2), pi/2, pi/2, pi/2).doit() == sqrt(6)/4
    assert Rotation.D(
        Rational(3, 2), Rational(-3, 2), Rational(3, 2), pi/2, pi/2, pi/2).doit() == sqrt(2)/4
    assert Rotation.D(
        Rational(3, 2), Rational(-3, 2), S.Half, pi/2, pi/2, pi/2).doit() == I*sqrt(6)/4
    assert Rotation.D(
        Rational(3, 2), Rational(-3, 2), Rational(-1, 2), pi/2, pi/2, pi/2).doit() == -sqrt(6)/4
    assert Rotation.D(
        Rational(3, 2), Rational(-3, 2), Rational(-3, 2), pi/2, pi/2, pi/2).doit() == -I*sqrt(2)/4
    # j = 2
    assert Rotation.D(2, 2, 2, pi/2, pi/2, pi/2).doit() == Rational(1, 4)
    assert Rotation.D(2, 2, 1, pi/2, pi/2, pi/2).doit() == -I/2
    assert Rotation.D(2, 2, 0, pi/2, pi/2, pi/2).doit() == -sqrt(6)/4
    assert Rotation.D(2, 2, -1, pi/2, pi/2, pi/2).doit() == I/2
    assert Rotation.D(2, 2, -2, pi/2, pi/2, pi/2).doit() == Rational(1, 4)
    assert Rotation.D(2, 1, 2, pi/2, pi/2, pi/2).doit() == I/2
    assert Rotation.D(2, 1, 1, pi/2, pi/2, pi/2).doit() == S.Half
    assert Rotation.D(2, 1, 0, pi/2, pi/2, pi/2).doit() == 0
    assert Rotation.D(2, 1, -1, pi/2, pi/2, pi/2).doit() == S.Half
    assert Rotation.D(2, 1, -2, pi/2, pi/2, pi/2).doit() == -I/2
    assert Rotation.D(2, 0, 2, pi/2, pi/2, pi/2).doit() == -sqrt(6)/4
    assert Rotation.D(2, 0, 1, pi/2, pi/2, pi/2).doit() == 0
    assert Rotation.D(2, 0, 0, pi/2, pi/2, pi/2).doit() == Rational(-1, 2)
    assert Rotation.D(2, 0, -1, pi/2, pi/2, pi/2).doit() == 0
    assert Rotation.D(2, 0, -2, pi/2, pi/2, pi/2).doit() == -sqrt(6)/4
    assert Rotation.D(2, -1, 2, pi/2, pi/2, pi/2).doit() == -I/2
    assert Rotation.D(2, -1, 1, pi/2, pi/2, pi/2).doit() == S.Half
    assert Rotation.D(2, -1, 0, pi/2, pi/2, pi/2).doit() == 0
    assert Rotation.D(2, -1, -1, pi/2, pi/2, pi/2).doit() == S.Half
    assert Rotation.D(2, -1, -2, pi/2, pi/2, pi/2).doit() == I/2
    assert Rotation.D(2, -2, 2, pi/2, pi/2, pi/2).doit() == Rational(1, 4)
    assert Rotation.D(2, -2, 1, pi/2, pi/2, pi/2).doit() == I/2
    assert Rotation.D(2, -2, 0, pi/2, pi/2, pi/2).doit() == -sqrt(6)/4
    assert Rotation.D(2, -2, -1, pi/2, pi/2, pi/2).doit() == -I/2
    assert Rotation.D(2, -2, -2, pi/2, pi/2, pi/2).doit() == Rational(1, 4)


def test_wignerd():
    assert Rotation.D(
        j, m, mp, alpha, beta, gamma) == WignerD(j, m, mp, alpha, beta, gamma)
    assert Rotation.d(j, m, mp, beta) == WignerD(j, m, mp, 0, beta, 0)

def test_wignerD():
    i,j=symbols('i j')
    assert Rotation.D(1, 1, 1, 0, 0, 0) == WignerD(1, 1, 1, 0, 0, 0)
    assert Rotation.D(1, 1, 2, 0, 0, 0) == WignerD(1, 1, 2, 0, 0, 0)
    assert Rotation.D(1, i**2 - j**2, i**2 - j**2, 0, 0, 0) == WignerD(1, i**2 - j**2, i**2 - j**2, 0, 0, 0)
    assert Rotation.D(1, i, i, 0, 0, 0) == WignerD(1, i, i, 0, 0, 0)
    assert Rotation.D(1, i, i+1, 0, 0, 0) == WignerD(1, i, i+1, 0, 0, 0)
    assert Rotation.D(1, 0, 0, 0, 0, 0) == WignerD(1, 0, 0, 0, 0, 0)

def test_jplus():
    assert Commutator(Jplus, Jminus).doit() == 2*hbar*Jz
    assert Jplus.matrix_element(1, 1, 1, 1) == 0
    assert Jplus.rewrite('xyz') == Jx + I*Jy
    # Normal operators, normal states
    # Numerical
    assert qapply(Jplus*JxKet(1, 1)) == \
        -hbar*sqrt(2)*JxKet(1, 0)/2 + hbar*JxKet(1, 1)
    assert qapply(Jplus*JyKet(1, 1)) == \
        hbar*sqrt(2)*JyKet(1, 0)/2 + I*hbar*JyKet(1, 1)
    assert qapply(Jplus*JzKet(1, 1)) == 0
    # Symbolic
    assert qapply(Jplus*JxKet(j, m)) == \
        Sum(hbar * sqrt(-mi**2 - mi + j**2 + j) * WignerD(j, mi, m, 0, pi/2, 0) *
        Sum(WignerD(j, mi1, mi + 1, 0, pi*Rational(3, 2), 0) * JxKet(j, mi1),
        (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jplus*JyKet(j, m)) == \
        Sum(hbar * sqrt(j**2 + j - mi**2 - mi) * WignerD(j, mi, m, pi*Rational(3, 2), -pi/2, pi/2) *
        Sum(WignerD(j, mi1, mi + 1, pi*Rational(3, 2), pi/2, pi/2) * JyKet(j, mi1),
        (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jplus*JzKet(j, m)) == \
        hbar*sqrt(j**2 + j - m**2 - m)*JzKet(j, m + 1)
    # Normal operators, coupled states
    # Numerical
    assert qapply(Jplus*JxKetCoupled(1, 1, (1, 1))) == -hbar*sqrt(2) * \
        JxKetCoupled(1, 0, (1, 1))/2 + hbar*JxKetCoupled(1, 1, (1, 1))
    assert qapply(Jplus*JyKetCoupled(1, 1, (1, 1))) == hbar*sqrt(2) * \
        JyKetCoupled(1, 0, (1, 1))/2 + I*hbar*JyKetCoupled(1, 1, (1, 1))
    assert qapply(Jplus*JzKet(1, 1)) == 0
    # Symbolic
    assert qapply(Jplus*JxKetCoupled(j, m, (j1, j2))) == \
        Sum(hbar * sqrt(-mi**2 - mi + j**2 + j) * WignerD(j, mi, m, 0, pi/2, 0) *
        Sum(
            WignerD(
                j, mi1, mi + 1, 0, pi*Rational(3, 2), 0) * JxKetCoupled(j, mi1, (j1, j2)),
        (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jplus*JyKetCoupled(j, m, (j1, j2))) == \
        Sum(hbar * sqrt(j**2 + j - mi**2 - mi) * WignerD(j, mi, m, pi*Rational(3, 2), -pi/2, pi/2) *
        Sum(
            WignerD(j, mi1, mi + 1, pi*Rational(3, 2), pi/2, pi/2) *
            JyKetCoupled(j, mi1, (j1, j2)),
        (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jplus*JzKetCoupled(j, m, (j1, j2))) == \
        hbar*sqrt(j**2 + j - m**2 - m)*JzKetCoupled(j, m + 1, (j1, j2))
    # Uncoupled operators, uncoupled states
    # Numerical
    e1 = qapply(TensorProduct(Jplus, 1)*TensorProduct(JxKet(1, 1), JxKet(1, -1)))
    e2 = -hbar*sqrt(2)*TensorProduct(JxKet(1, 0), JxKet(1, -1))/2 + \
        hbar*TensorProduct(JxKet(1, 1), JxKet(1, -1))
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(1, Jplus)*TensorProduct(JxKet(1, 1), JxKet(1, -1)))
    e2 = -hbar*TensorProduct(JxKet(1, 1), JxKet(1, -1)) + \
        hbar*sqrt(2)*TensorProduct(JxKet(1, 1), JxKet(1, 0))/2
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(Jplus, 1)*TensorProduct(JyKet(1, 1), JyKet(1, -1)))
    e2 = hbar*sqrt(2)*TensorProduct(JyKet(1, 0), JyKet(1, -1))/2 + \
        hbar*I*TensorProduct(JyKet(1, 1), JyKet(1, -1))
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(1, Jplus)*TensorProduct(JyKet(1, 1), JyKet(1, -1)))
    e2 = -hbar*I*TensorProduct(JyKet(1, 1), JyKet(1, -1)) + \
        hbar*sqrt(2)*TensorProduct(JyKet(1, 1), JyKet(1, 0))/2
    assert_simplify_expand(e1, e2)
    assert qapply(
        TensorProduct(Jplus, 1)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == 0
    assert qapply(TensorProduct(1, Jplus)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        hbar*sqrt(2)*TensorProduct(JzKet(1, 1), JzKet(1, 0))
    # Symbolic
    assert qapply(TensorProduct(Jplus, 1)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        TensorProduct(Sum(hbar * sqrt(-mi**2 - mi + j1**2 + j1) * WignerD(j1, mi, m1, 0, pi/2, 0) *
        Sum(WignerD(j1, mi1, mi + 1, 0, pi*Rational(3, 2), 0) * JxKet(j1, mi1),
        (mi1, -j1, j1)), (mi, -j1, j1)), JxKet(j2, m2))
    assert qapply(TensorProduct(1, Jplus)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        TensorProduct(JxKet(j1, m1), Sum(hbar * sqrt(-mi**2 - mi + j2**2 + j2) * WignerD(j2, mi, m2, 0, pi/2, 0) *
        Sum(WignerD(j2, mi1, mi + 1, 0, pi*Rational(3, 2), 0) * JxKet(j2, mi1),
        (mi1, -j2, j2)), (mi, -j2, j2)))
    assert qapply(TensorProduct(Jplus, 1)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        TensorProduct(Sum(hbar * sqrt(j1**2 + j1 - mi**2 - mi) * WignerD(j1, mi, m1, pi*Rational(3, 2), -pi/2, pi/2) *
        Sum(WignerD(j1, mi1, mi + 1, pi*Rational(3, 2), pi/2, pi/2) * JyKet(j1, mi1),
        (mi1, -j1, j1)), (mi, -j1, j1)), JyKet(j2, m2))
    assert qapply(TensorProduct(1, Jplus)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        TensorProduct(JyKet(j1, m1), Sum(hbar * sqrt(j2**2 + j2 - mi**2 - mi) * WignerD(j2, mi, m2, pi*Rational(3, 2), -pi/2, pi/2) *
        Sum(WignerD(j2, mi1, mi + 1, pi*Rational(3, 2), pi/2, pi/2) * JyKet(j2, mi1),
        (mi1, -j2, j2)), (mi, -j2, j2)))
    assert qapply(TensorProduct(Jplus, 1)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))) == \
        hbar*sqrt(
            j1**2 + j1 - m1**2 - m1)*TensorProduct(JzKet(j1, m1 + 1), JzKet(j2, m2))
    assert qapply(TensorProduct(1, Jplus)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))) == \
        hbar*sqrt(
            j2**2 + j2 - m2**2 - m2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2 + 1))


def test_jminus():
    assert qapply(Jminus*JzKet(1, -1)) == 0
    assert Jminus.matrix_element(1, 0, 1, 1) == sqrt(2)*hbar
    assert Jminus.rewrite('xyz') == Jx - I*Jy
    # Normal operators, normal states
    # Numerical
    assert qapply(Jminus*JxKet(1, 1)) == \
        hbar*sqrt(2)*JxKet(1, 0)/2 + hbar*JxKet(1, 1)
    assert qapply(Jminus*JyKet(1, 1)) == \
        hbar*sqrt(2)*JyKet(1, 0)/2 - hbar*I*JyKet(1, 1)
    assert qapply(Jminus*JzKet(1, 1)) == sqrt(2)*hbar*JzKet(1, 0)
    # Symbolic
    assert qapply(Jminus*JxKet(j, m)) == \
        Sum(hbar*sqrt(j**2 + j - mi**2 + mi)*WignerD(j, mi, m, 0, pi/2, 0) *
        Sum(WignerD(j, mi1, mi - 1, 0, pi*Rational(3, 2), 0)*JxKet(j, mi1),
        (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jminus*JyKet(j, m)) == \
        Sum(hbar*sqrt(j**2 + j - mi**2 + mi)*WignerD(j, mi, m, pi*Rational(3, 2), -pi/2, pi/2) *
        Sum(WignerD(j, mi1, mi - 1, pi*Rational(3, 2), pi/2, pi/2)*JyKet(j, mi1),
        (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jminus*JzKet(j, m)) == \
        hbar*sqrt(j**2 + j - m**2 + m)*JzKet(j, m - 1)
    # Normal operators, coupled states
    # Numerical
    assert qapply(Jminus*JxKetCoupled(1, 1, (1, 1))) == \
        hbar*sqrt(2)*JxKetCoupled(1, 0, (1, 1))/2 + \
        hbar*JxKetCoupled(1, 1, (1, 1))
    assert qapply(Jminus*JyKetCoupled(1, 1, (1, 1))) == \
        hbar*sqrt(2)*JyKetCoupled(1, 0, (1, 1))/2 - \
        hbar*I*JyKetCoupled(1, 1, (1, 1))
    assert qapply(Jminus*JzKetCoupled(1, 1, (1, 1))) == \
        sqrt(2)*hbar*JzKetCoupled(1, 0, (1, 1))
    # Symbolic
    assert qapply(Jminus*JxKetCoupled(j, m, (j1, j2))) == \
        Sum(hbar*sqrt(j**2 + j - mi**2 + mi)*WignerD(j, mi, m, 0, pi/2, 0) *
        Sum(WignerD(j, mi1, mi - 1, 0, pi*Rational(3, 2), 0)*JxKetCoupled(j, mi1, (j1, j2)),
        (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jminus*JyKetCoupled(j, m, (j1, j2))) == \
        Sum(hbar*sqrt(j**2 + j - mi**2 + mi)*WignerD(j, mi, m, pi*Rational(3, 2), -pi/2, pi/2) *
        Sum(
            WignerD(j, mi1, mi - 1, pi*Rational(3, 2), pi/2, pi/2)*
            JyKetCoupled(j, mi1, (j1, j2)),
        (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jminus*JzKetCoupled(j, m, (j1, j2))) == \
        hbar*sqrt(j**2 + j - m**2 + m)*JzKetCoupled(j, m - 1, (j1, j2))
    # Uncoupled operators, uncoupled states
    # Numerical
    e1 = qapply(TensorProduct(Jminus, 1)*TensorProduct(JxKet(1, 1), JxKet(1, -1)))
    e2 = hbar*sqrt(2)*TensorProduct(JxKet(1, 0), JxKet(1, -1))/2 + \
        hbar*TensorProduct(JxKet(1, 1), JxKet(1, -1))
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(1, Jminus)*TensorProduct(JxKet(1, 1), JxKet(1, -1)))
    e2 = -hbar*TensorProduct(JxKet(1, 1), JxKet(1, -1)) - \
        hbar*sqrt(2)*TensorProduct(JxKet(1, 1), JxKet(1, 0))/2
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(Jminus, 1)*TensorProduct(JyKet(1, 1), JyKet(1, -1)))
    e2 = hbar*sqrt(2)*TensorProduct(JyKet(1, 0), JyKet(1, -1))/2 - \
        hbar*I*TensorProduct(JyKet(1, 1), JyKet(1, -1))
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(1, Jminus)*TensorProduct(JyKet(1, 1), JyKet(1, -1)))
    e2 = hbar*I*TensorProduct(JyKet(1, 1), JyKet(1, -1)) + \
        hbar*sqrt(2)*TensorProduct(JyKet(1, 1), JyKet(1, 0))/2
    assert_simplify_expand(e1, e2)
    assert qapply(TensorProduct(Jminus, 1)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        sqrt(2)*hbar*TensorProduct(JzKet(1, 0), JzKet(1, -1))
    assert qapply(TensorProduct(
        1, Jminus)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == 0
    # Symbolic
    assert qapply(TensorProduct(Jminus, 1)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        TensorProduct(Sum(hbar*sqrt(j1**2 + j1 - mi**2 + mi)*WignerD(j1, mi, m1, 0, pi/2, 0) *
        Sum(WignerD(j1, mi1, mi - 1, 0, pi*Rational(3, 2), 0)*JxKet(j1, mi1),
        (mi1, -j1, j1)), (mi, -j1, j1)), JxKet(j2, m2))
    assert qapply(TensorProduct(1, Jminus)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        TensorProduct(JxKet(j1, m1), Sum(hbar*sqrt(j2**2 + j2 - mi**2 + mi)*WignerD(j2, mi, m2, 0, pi/2, 0) *
        Sum(WignerD(j2, mi1, mi - 1, 0, pi*Rational(3, 2), 0)*JxKet(j2, mi1),
        (mi1, -j2, j2)), (mi, -j2, j2)))
    assert qapply(TensorProduct(Jminus, 1)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        TensorProduct(Sum(hbar*sqrt(j1**2 + j1 - mi**2 + mi)*WignerD(j1, mi, m1, pi*Rational(3, 2), -pi/2, pi/2) *
        Sum(WignerD(j1, mi1, mi - 1, pi*Rational(3, 2), pi/2, pi/2)*JyKet(j1, mi1),
        (mi1, -j1, j1)), (mi, -j1, j1)), JyKet(j2, m2))
    assert qapply(TensorProduct(1, Jminus)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        TensorProduct(JyKet(j1, m1), Sum(hbar*sqrt(j2**2 + j2 - mi**2 + mi)*WignerD(j2, mi, m2, pi*Rational(3, 2), -pi/2, pi/2) *
        Sum(WignerD(j2, mi1, mi - 1, pi*Rational(3, 2), pi/2, pi/2)*JyKet(j2, mi1),
        (mi1, -j2, j2)), (mi, -j2, j2)))
    assert qapply(TensorProduct(Jminus, 1)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))) == \
        hbar*sqrt(
            j1**2 + j1 - m1**2 + m1)*TensorProduct(JzKet(j1, m1 - 1), JzKet(j2, m2))
    assert qapply(TensorProduct(1, Jminus)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))) == \
        hbar*sqrt(
            j2**2 + j2 - m2**2 + m2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2 - 1))


def test_j2():
    assert Commutator(J2, Jz).doit() == 0
    assert J2.matrix_element(1, 1, 1, 1) == 2*hbar**2
    # Normal operators, normal states
    # Numerical
    assert qapply(J2*JxKet(1, 1)) == 2*hbar**2*JxKet(1, 1)
    assert qapply(J2*JyKet(1, 1)) == 2*hbar**2*JyKet(1, 1)
    assert qapply(J2*JzKet(1, 1)) == 2*hbar**2*JzKet(1, 1)
    # Symbolic
    assert qapply(J2*JxKet(j, m)) == \
        hbar**2*j**2*JxKet(j, m) + hbar**2*j*JxKet(j, m)
    assert qapply(J2*JyKet(j, m)) == \
        hbar**2*j**2*JyKet(j, m) + hbar**2*j*JyKet(j, m)
    assert qapply(J2*JzKet(j, m)) == \
        hbar**2*j**2*JzKet(j, m) + hbar**2*j*JzKet(j, m)
    # Normal operators, coupled states
    # Numerical
    assert qapply(J2*JxKetCoupled(1, 1, (1, 1))) == \
        2*hbar**2*JxKetCoupled(1, 1, (1, 1))
    assert qapply(J2*JyKetCoupled(1, 1, (1, 1))) == \
        2*hbar**2*JyKetCoupled(1, 1, (1, 1))
    assert qapply(J2*JzKetCoupled(1, 1, (1, 1))) == \
        2*hbar**2*JzKetCoupled(1, 1, (1, 1))
    # Symbolic
    assert qapply(J2*JxKetCoupled(j, m, (j1, j2))) == \
        hbar**2*j**2*JxKetCoupled(j, m, (j1, j2)) + \
        hbar**2*j*JxKetCoupled(j, m, (j1, j2))
    assert qapply(J2*JyKetCoupled(j, m, (j1, j2))) == \
        hbar**2*j**2*JyKetCoupled(j, m, (j1, j2)) + \
        hbar**2*j*JyKetCoupled(j, m, (j1, j2))
    assert qapply(J2*JzKetCoupled(j, m, (j1, j2))) == \
        hbar**2*j**2*JzKetCoupled(j, m, (j1, j2)) + \
        hbar**2*j*JzKetCoupled(j, m, (j1, j2))
    # Uncoupled operators, uncoupled states
    # Numerical
    assert qapply(TensorProduct(J2, 1)*TensorProduct(JxKet(1, 1), JxKet(1, -1))) == \
        2*hbar**2*TensorProduct(JxKet(1, 1), JxKet(1, -1))
    assert qapply(TensorProduct(1, J2)*TensorProduct(JxKet(1, 1), JxKet(1, -1))) == \
        2*hbar**2*TensorProduct(JxKet(1, 1), JxKet(1, -1))
    assert qapply(TensorProduct(J2, 1)*TensorProduct(JyKet(1, 1), JyKet(1, -1))) == \
        2*hbar**2*TensorProduct(JyKet(1, 1), JyKet(1, -1))
    assert qapply(TensorProduct(1, J2)*TensorProduct(JyKet(1, 1), JyKet(1, -1))) == \
        2*hbar**2*TensorProduct(JyKet(1, 1), JyKet(1, -1))
    assert qapply(TensorProduct(J2, 1)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        2*hbar**2*TensorProduct(JzKet(1, 1), JzKet(1, -1))
    assert qapply(TensorProduct(1, J2)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        2*hbar**2*TensorProduct(JzKet(1, 1), JzKet(1, -1))
    # Symbolic
    e1 = qapply(TensorProduct(J2, 1)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2)))
    e2 = hbar**2*j1**2*TensorProduct(JxKet(j1, m1), JxKet(j2, m2)) + \
        hbar**2*j1*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(1, J2)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2)))
    e2 = hbar**2*j2**2*TensorProduct(JxKet(j1, m1), JxKet(j2, m2)) + \
        hbar**2*j2*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(J2, 1)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2)))
    e2 = hbar**2*j1**2*TensorProduct(JyKet(j1, m1), JyKet(j2, m2)) + \
        hbar**2*j1*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(1, J2)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2)))
    e2 = hbar**2*j2**2*TensorProduct(JyKet(j1, m1), JyKet(j2, m2)) + \
        hbar**2*j2*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(J2, 1)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2)))
    e2 = hbar**2*j1**2*TensorProduct(JzKet(j1, m1), JzKet(j2, m2)) + \
        hbar**2*j1*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(1, J2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2)))
    e2 = hbar**2*j2**2*TensorProduct(JzKet(j1, m1), JzKet(j2, m2)) + \
        hbar**2*j2*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))
    assert_simplify_expand(e1, e2)


def test_jx():
    assert Commutator(Jx, Jz).doit() == -I*hbar*Jy
    assert Jx.rewrite('plusminus') == (Jminus + Jplus)/2
    assert represent(Jx, basis=Jz, j=1) == (
        represent(Jplus, basis=Jz, j=1) + represent(Jminus, basis=Jz, j=1))/2
    # Normal operators, normal states
    # Numerical
    assert qapply(Jx*JxKet(1, 1)) == hbar*JxKet(1, 1)
    assert qapply(Jx*JyKet(1, 1)) == hbar*JyKet(1, 1)
    assert qapply(Jx*JzKet(1, 1)) == sqrt(2)*hbar*JzKet(1, 0)/2
    # Symbolic
    assert qapply(Jx*JxKet(j, m)) == hbar*m*JxKet(j, m)
    assert qapply(Jx*JyKet(j, m)) == \
        Sum(hbar*mi*WignerD(j, mi, m, 0, 0, pi/2)*Sum(WignerD(j,
            mi1, mi, pi*Rational(3, 2), 0, 0)*JyKet(j, mi1), (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jx*JzKet(j, m)) == \
        hbar*sqrt(j**2 + j - m**2 - m)*JzKet(j, m + 1)/2 + hbar*sqrt(j**2 +
                  j - m**2 + m)*JzKet(j, m - 1)/2
    # Normal operators, coupled states
    # Numerical
    assert qapply(Jx*JxKetCoupled(1, 1, (1, 1))) == \
        hbar*JxKetCoupled(1, 1, (1, 1))
    assert qapply(Jx*JyKetCoupled(1, 1, (1, 1))) == \
        hbar*JyKetCoupled(1, 1, (1, 1))
    assert qapply(Jx*JzKetCoupled(1, 1, (1, 1))) == \
        sqrt(2)*hbar*JzKetCoupled(1, 0, (1, 1))/2
    # Symbolic
    assert qapply(Jx*JxKetCoupled(j, m, (j1, j2))) == \
        hbar*m*JxKetCoupled(j, m, (j1, j2))
    assert qapply(Jx*JyKetCoupled(j, m, (j1, j2))) == \
        Sum(hbar*mi*WignerD(j, mi, m, 0, 0, pi/2)*Sum(WignerD(j, mi1, mi, pi*Rational(3, 2), 0, 0)*JyKetCoupled(j, mi1, (j1, j2)), (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jx*JzKetCoupled(j, m, (j1, j2))) == \
        hbar*sqrt(j**2 + j - m**2 - m)*JzKetCoupled(j, m + 1, (j1, j2))/2 + \
        hbar*sqrt(j**2 + j - m**2 + m)*JzKetCoupled(j, m - 1, (j1, j2))/2
    # Normal operators, uncoupled states
    # Numerical
    assert qapply(Jx*TensorProduct(JxKet(1, 1), JxKet(1, 1))) == \
        2*hbar*TensorProduct(JxKet(1, 1), JxKet(1, 1))
    assert qapply(Jx*TensorProduct(JyKet(1, 1), JyKet(1, 1))) == \
        hbar*TensorProduct(JyKet(1, 1), JyKet(1, 1)) + \
        hbar*TensorProduct(JyKet(1, 1), JyKet(1, 1))
    assert qapply(Jx*TensorProduct(JzKet(1, 1), JzKet(1, 1))) == \
        sqrt(2)*hbar*TensorProduct(JzKet(1, 1), JzKet(1, 0))/2 + \
        sqrt(2)*hbar*TensorProduct(JzKet(1, 0), JzKet(1, 1))/2
    assert qapply(Jx*TensorProduct(JxKet(1, 1), JxKet(1, -1))) == 0
    # Symbolic
    assert qapply(Jx*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        hbar*m1*TensorProduct(JxKet(j1, m1), JxKet(j2, m2)) + \
        hbar*m2*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))
    assert qapply(Jx*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        TensorProduct(Sum(hbar*mi*WignerD(j1, mi, m1, 0, 0, pi/2)*Sum(WignerD(j1, mi1, mi, pi*Rational(3, 2), 0, 0)*JyKet(j1, mi1), (mi1, -j1, j1)), (mi, -j1, j1)), JyKet(j2, m2)) + \
        TensorProduct(JyKet(j1, m1), Sum(hbar*mi*WignerD(j2, mi, m2, 0, 0, pi/2)*Sum(WignerD(j2, mi1, mi, pi*Rational(3, 2), 0, 0)*JyKet(j2, mi1), (mi1, -j2, j2)), (mi, -j2, j2)))
    assert qapply(Jx*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))) == \
        hbar*sqrt(j1**2 + j1 - m1**2 - m1)*TensorProduct(JzKet(j1, m1 + 1), JzKet(j2, m2))/2 + \
        hbar*sqrt(j1**2 + j1 - m1**2 + m1)*TensorProduct(JzKet(j1, m1 - 1), JzKet(j2, m2))/2 + \
        hbar*sqrt(j2**2 + j2 - m2**2 - m2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2 + 1))/2 + \
        hbar*sqrt(
            j2**2 + j2 - m2**2 + m2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2 - 1))/2
    # Uncoupled operators, uncoupled states
    # Numerical
    assert qapply(TensorProduct(Jx, 1)*TensorProduct(JxKet(1, 1), JxKet(1, -1))) == \
        hbar*TensorProduct(JxKet(1, 1), JxKet(1, -1))
    assert qapply(TensorProduct(1, Jx)*TensorProduct(JxKet(1, 1), JxKet(1, -1))) == \
        -hbar*TensorProduct(JxKet(1, 1), JxKet(1, -1))
    assert qapply(TensorProduct(Jx, 1)*TensorProduct(JyKet(1, 1), JyKet(1, -1))) == \
        hbar*TensorProduct(JyKet(1, 1), JyKet(1, -1))
    assert qapply(TensorProduct(1, Jx)*TensorProduct(JyKet(1, 1), JyKet(1, -1))) == \
        -hbar*TensorProduct(JyKet(1, 1), JyKet(1, -1))
    assert qapply(TensorProduct(Jx, 1)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        hbar*sqrt(2)*TensorProduct(JzKet(1, 0), JzKet(1, -1))/2
    assert qapply(TensorProduct(1, Jx)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        hbar*sqrt(2)*TensorProduct(JzKet(1, 1), JzKet(1, 0))/2
    # Symbolic
    assert qapply(TensorProduct(Jx, 1)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        hbar*m1*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))
    assert qapply(TensorProduct(1, Jx)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        hbar*m2*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))
    assert qapply(TensorProduct(Jx, 1)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        TensorProduct(Sum(hbar*mi*WignerD(j1, mi, m1, 0, 0, pi/2) * Sum(WignerD(j1, mi1, mi, pi*Rational(3, 2), 0, 0)*JyKet(j1, mi1), (mi1, -j1, j1)), (mi, -j1, j1)), JyKet(j2, m2))
    assert qapply(TensorProduct(1, Jx)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        TensorProduct(JyKet(j1, m1), Sum(hbar*mi*WignerD(j2, mi, m2, 0, 0, pi/2) * Sum(WignerD(j2, mi1, mi, pi*Rational(3, 2), 0, 0)*JyKet(j2, mi1), (mi1, -j2, j2)), (mi, -j2, j2)))
    e1 = qapply(TensorProduct(Jx, 1)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2)))
    e2 = hbar*sqrt(j1**2 + j1 - m1**2 - m1)*TensorProduct(JzKet(j1, m1 + 1), JzKet(j2, m2))/2 + \
        hbar*sqrt(
            j1**2 + j1 - m1**2 + m1)*TensorProduct(JzKet(j1, m1 - 1), JzKet(j2, m2))/2
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(1, Jx)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2)))
    e2 = hbar*sqrt(j2**2 + j2 - m2**2 - m2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2 + 1))/2 + \
        hbar*sqrt(
            j2**2 + j2 - m2**2 + m2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2 - 1))/2
    assert_simplify_expand(e1, e2)


def test_jy():
    assert Commutator(Jy, Jz).doit() == I*hbar*Jx
    assert Jy.rewrite('plusminus') == (Jplus - Jminus)/(2*I)
    assert represent(Jy, basis=Jz) == (
        represent(Jplus, basis=Jz) - represent(Jminus, basis=Jz))/(2*I)
    # Normal operators, normal states
    # Numerical
    assert qapply(Jy*JxKet(1, 1)) == hbar*JxKet(1, 1)
    assert qapply(Jy*JyKet(1, 1)) == hbar*JyKet(1, 1)
    assert qapply(Jy*JzKet(1, 1)) == sqrt(2)*hbar*I*JzKet(1, 0)/2
    # Symbolic
    assert qapply(Jy*JxKet(j, m)) == \
        Sum(hbar*mi*WignerD(j, mi, m, pi*Rational(3, 2), 0, 0)*Sum(WignerD(
            j, mi1, mi, 0, 0, pi/2)*JxKet(j, mi1), (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jy*JyKet(j, m)) == hbar*m*JyKet(j, m)
    assert qapply(Jy*JzKet(j, m)) == \
        -hbar*I*sqrt(j**2 + j - m**2 - m)*JzKet(
            j, m + 1)/2 + hbar*I*sqrt(j**2 + j - m**2 + m)*JzKet(j, m - 1)/2
    # Normal operators, coupled states
    # Numerical
    assert qapply(Jy*JxKetCoupled(1, 1, (1, 1))) == \
        hbar*JxKetCoupled(1, 1, (1, 1))
    assert qapply(Jy*JyKetCoupled(1, 1, (1, 1))) == \
        hbar*JyKetCoupled(1, 1, (1, 1))
    assert qapply(Jy*JzKetCoupled(1, 1, (1, 1))) == \
        sqrt(2)*hbar*I*JzKetCoupled(1, 0, (1, 1))/2
    # Symbolic
    assert qapply(Jy*JxKetCoupled(j, m, (j1, j2))) == \
        Sum(hbar*mi*WignerD(j, mi, m, pi*Rational(3, 2), 0, 0)*Sum(WignerD(j, mi1, mi, 0, 0, pi/2)*JxKetCoupled(j, mi1, (j1, j2)), (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jy*JyKetCoupled(j, m, (j1, j2))) == \
        hbar*m*JyKetCoupled(j, m, (j1, j2))
    assert qapply(Jy*JzKetCoupled(j, m, (j1, j2))) == \
        -hbar*I*sqrt(j**2 + j - m**2 - m)*JzKetCoupled(j, m + 1, (j1, j2))/2 + \
        hbar*I*sqrt(j**2 + j - m**2 + m)*JzKetCoupled(j, m - 1, (j1, j2))/2
    # Normal operators, uncoupled states
    # Numerical
    assert qapply(Jy*TensorProduct(JxKet(1, 1), JxKet(1, 1))) == \
        hbar*TensorProduct(JxKet(1, 1), JxKet(1, 1)) + \
        hbar*TensorProduct(JxKet(1, 1), JxKet(1, 1))
    assert qapply(Jy*TensorProduct(JyKet(1, 1), JyKet(1, 1))) == \
        2*hbar*TensorProduct(JyKet(1, 1), JyKet(1, 1))
    assert qapply(Jy*TensorProduct(JzKet(1, 1), JzKet(1, 1))) == \
        sqrt(2)*hbar*I*TensorProduct(JzKet(1, 1), JzKet(1, 0))/2 + \
        sqrt(2)*hbar*I*TensorProduct(JzKet(1, 0), JzKet(1, 1))/2
    assert qapply(Jy*TensorProduct(JyKet(1, 1), JyKet(1, -1))) == 0
    # Symbolic
    assert qapply(Jy*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        TensorProduct(JxKet(j1, m1), Sum(hbar*mi*WignerD(j2, mi, m2, pi*Rational(3, 2), 0, 0)*Sum(WignerD(j2, mi1, mi, 0, 0, pi/2)*JxKet(j2, mi1), (mi1, -j2, j2)), (mi, -j2, j2))) + \
        TensorProduct(Sum(hbar*mi*WignerD(j1, mi, m1, pi*Rational(3, 2), 0, 0)*Sum(WignerD(j1, mi1, mi, 0, 0, pi/2)*JxKet(j1, mi1), (mi1, -j1, j1)), (mi, -j1, j1)), JxKet(j2, m2))
    assert qapply(Jy*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        hbar*m1*TensorProduct(JyKet(j1, m1), JyKet(
            j2, m2)) + hbar*m2*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))
    assert qapply(Jy*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))) == \
        -hbar*I*sqrt(j1**2 + j1 - m1**2 - m1)*TensorProduct(JzKet(j1, m1 + 1), JzKet(j2, m2))/2 + \
        hbar*I*sqrt(j1**2 + j1 - m1**2 + m1)*TensorProduct(JzKet(j1, m1 - 1), JzKet(j2, m2))/2 + \
        -hbar*I*sqrt(j2**2 + j2 - m2**2 - m2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2 + 1))/2 + \
        hbar*I*sqrt(
            j2**2 + j2 - m2**2 + m2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2 - 1))/2
    # Uncoupled operators, uncoupled states
    # Numerical
    assert qapply(TensorProduct(Jy, 1)*TensorProduct(JxKet(1, 1), JxKet(1, -1))) == \
        hbar*TensorProduct(JxKet(1, 1), JxKet(1, -1))
    assert qapply(TensorProduct(1, Jy)*TensorProduct(JxKet(1, 1), JxKet(1, -1))) == \
        -hbar*TensorProduct(JxKet(1, 1), JxKet(1, -1))
    assert qapply(TensorProduct(Jy, 1)*TensorProduct(JyKet(1, 1), JyKet(1, -1))) == \
        hbar*TensorProduct(JyKet(1, 1), JyKet(1, -1))
    assert qapply(TensorProduct(1, Jy)*TensorProduct(JyKet(1, 1), JyKet(1, -1))) == \
        -hbar*TensorProduct(JyKet(1, 1), JyKet(1, -1))
    assert qapply(TensorProduct(Jy, 1)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        hbar*sqrt(2)*I*TensorProduct(JzKet(1, 0), JzKet(1, -1))/2
    assert qapply(TensorProduct(1, Jy)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        -hbar*sqrt(2)*I*TensorProduct(JzKet(1, 1), JzKet(1, 0))/2
    # Symbolic
    assert qapply(TensorProduct(Jy, 1)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        TensorProduct(Sum(hbar*mi*WignerD(j1, mi, m1, pi*Rational(3, 2), 0, 0) * Sum(WignerD(j1, mi1, mi, 0, 0, pi/2)*JxKet(j1, mi1), (mi1, -j1, j1)), (mi, -j1, j1)), JxKet(j2, m2))
    assert qapply(TensorProduct(1, Jy)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        TensorProduct(JxKet(j1, m1), Sum(hbar*mi*WignerD(j2, mi, m2, pi*Rational(3, 2), 0, 0) * Sum(WignerD(j2, mi1, mi, 0, 0, pi/2)*JxKet(j2, mi1), (mi1, -j2, j2)), (mi, -j2, j2)))
    assert qapply(TensorProduct(Jy, 1)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        hbar*m1*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))
    assert qapply(TensorProduct(1, Jy)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        hbar*m2*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))
    e1 = qapply(TensorProduct(Jy, 1)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2)))
    e2 = -hbar*I*sqrt(j1**2 + j1 - m1**2 - m1)*TensorProduct(JzKet(j1, m1 + 1), JzKet(j2, m2))/2 + \
        hbar*I*sqrt(
            j1**2 + j1 - m1**2 + m1)*TensorProduct(JzKet(j1, m1 - 1), JzKet(j2, m2))/2
    assert_simplify_expand(e1, e2)
    e1 = qapply(TensorProduct(1, Jy)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2)))
    e2 = -hbar*I*sqrt(j2**2 + j2 - m2**2 - m2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2 + 1))/2 + \
        hbar*I*sqrt(
            j2**2 + j2 - m2**2 + m2)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2 - 1))/2
    assert_simplify_expand(e1, e2)


def test_jz():
    assert Commutator(Jz, Jminus).doit() == -hbar*Jminus
    # Normal operators, normal states
    # Numerical
    assert qapply(Jz*JxKet(1, 1)) == -sqrt(2)*hbar*JxKet(1, 0)/2
    assert qapply(Jz*JyKet(1, 1)) == -sqrt(2)*hbar*I*JyKet(1, 0)/2
    assert qapply(Jz*JzKet(2, 1)) == hbar*JzKet(2, 1)
    # Symbolic
    assert qapply(Jz*JxKet(j, m)) == \
        Sum(hbar*mi*WignerD(j, mi, m, 0, pi/2, 0)*Sum(WignerD(j,
            mi1, mi, 0, pi*Rational(3, 2), 0)*JxKet(j, mi1), (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jz*JyKet(j, m)) == \
        Sum(hbar*mi*WignerD(j, mi, m, pi*Rational(3, 2), -pi/2, pi/2)*Sum(WignerD(j, mi1,
            mi, pi*Rational(3, 2), pi/2, pi/2)*JyKet(j, mi1), (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jz*JzKet(j, m)) == hbar*m*JzKet(j, m)
    # Normal operators, coupled states
    # Numerical
    assert qapply(Jz*JxKetCoupled(1, 1, (1, 1))) == \
        -sqrt(2)*hbar*JxKetCoupled(1, 0, (1, 1))/2
    assert qapply(Jz*JyKetCoupled(1, 1, (1, 1))) == \
        -sqrt(2)*hbar*I*JyKetCoupled(1, 0, (1, 1))/2
    assert qapply(Jz*JzKetCoupled(1, 1, (1, 1))) == \
        hbar*JzKetCoupled(1, 1, (1, 1))
    # Symbolic
    assert qapply(Jz*JxKetCoupled(j, m, (j1, j2))) == \
        Sum(hbar*mi*WignerD(j, mi, m, 0, pi/2, 0)*Sum(WignerD(j, mi1, mi, 0, pi*Rational(3, 2), 0)*JxKetCoupled(j, mi1, (j1, j2)), (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jz*JyKetCoupled(j, m, (j1, j2))) == \
        Sum(hbar*mi*WignerD(j, mi, m, pi*Rational(3, 2), -pi/2, pi/2)*Sum(WignerD(j, mi1, mi, pi*Rational(3, 2), pi/2, pi/2)*JyKetCoupled(j, mi1, (j1, j2)), (mi1, -j, j)), (mi, -j, j))
    assert qapply(Jz*JzKetCoupled(j, m, (j1, j2))) == \
        hbar*m*JzKetCoupled(j, m, (j1, j2))
    # Normal operators, uncoupled states
    # Numerical
    assert qapply(Jz*TensorProduct(JxKet(1, 1), JxKet(1, 1))) == \
        -sqrt(2)*hbar*TensorProduct(JxKet(1, 1), JxKet(1, 0))/2 - \
        sqrt(2)*hbar*TensorProduct(JxKet(1, 0), JxKet(1, 1))/2
    assert qapply(Jz*TensorProduct(JyKet(1, 1), JyKet(1, 1))) == \
        -sqrt(2)*hbar*I*TensorProduct(JyKet(1, 1), JyKet(1, 0))/2 - \
        sqrt(2)*hbar*I*TensorProduct(JyKet(1, 0), JyKet(1, 1))/2
    assert qapply(Jz*TensorProduct(JzKet(1, 1), JzKet(1, 1))) == \
        2*hbar*TensorProduct(JzKet(1, 1), JzKet(1, 1))
    assert qapply(Jz*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == 0
    # Symbolic
    assert qapply(Jz*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        TensorProduct(JxKet(j1, m1), Sum(hbar*mi*WignerD(j2, mi, m2, 0, pi/2, 0)*Sum(WignerD(j2, mi1, mi, 0, pi*Rational(3, 2), 0)*JxKet(j2, mi1), (mi1, -j2, j2)), (mi, -j2, j2))) + \
        TensorProduct(Sum(hbar*mi*WignerD(j1, mi, m1, 0, pi/2, 0)*Sum(WignerD(j1, mi1, mi, 0, pi*Rational(3, 2), 0)*JxKet(j1, mi1), (mi1, -j1, j1)), (mi, -j1, j1)), JxKet(j2, m2))
    assert qapply(Jz*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        TensorProduct(JyKet(j1, m1), Sum(hbar*mi*WignerD(j2, mi, m2, pi*Rational(3, 2), -pi/2, pi/2)*Sum(WignerD(j2, mi1, mi, pi*Rational(3, 2), pi/2, pi/2)*JyKet(j2, mi1), (mi1, -j2, j2)), (mi, -j2, j2))) + \
        TensorProduct(Sum(hbar*mi*WignerD(j1, mi, m1, pi*Rational(3, 2), -pi/2, pi/2)*Sum(WignerD(j1, mi1, mi, pi*Rational(3, 2), pi/2, pi/2)*JyKet(j1, mi1), (mi1, -j1, j1)), (mi, -j1, j1)), JyKet(j2, m2))
    assert qapply(Jz*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))) == \
        hbar*m1*TensorProduct(JzKet(j1, m1), JzKet(
            j2, m2)) + hbar*m2*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))
    # Uncoupled Operators
    # Numerical
    assert qapply(TensorProduct(Jz, 1)*TensorProduct(JxKet(1, 1), JxKet(1, -1))) == \
        -sqrt(2)*hbar*TensorProduct(JxKet(1, 0), JxKet(1, -1))/2
    assert qapply(TensorProduct(1, Jz)*TensorProduct(JxKet(1, 1), JxKet(1, -1))) == \
        -sqrt(2)*hbar*TensorProduct(JxKet(1, 1), JxKet(1, 0))/2
    assert qapply(TensorProduct(Jz, 1)*TensorProduct(JyKet(1, 1), JyKet(1, -1))) == \
        -sqrt(2)*I*hbar*TensorProduct(JyKet(1, 0), JyKet(1, -1))/2
    assert qapply(TensorProduct(1, Jz)*TensorProduct(JyKet(1, 1), JyKet(1, -1))) == \
        sqrt(2)*I*hbar*TensorProduct(JyKet(1, 1), JyKet(1, 0))/2
    assert qapply(TensorProduct(Jz, 1)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        hbar*TensorProduct(JzKet(1, 1), JzKet(1, -1))
    assert qapply(TensorProduct(1, Jz)*TensorProduct(JzKet(1, 1), JzKet(1, -1))) == \
        -hbar*TensorProduct(JzKet(1, 1), JzKet(1, -1))
    # Symbolic
    assert qapply(TensorProduct(Jz, 1)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        TensorProduct(Sum(hbar*mi*WignerD(j1, mi, m1, 0, pi/2, 0)*Sum(WignerD(j1, mi1, mi, 0, pi*Rational(3, 2), 0)*JxKet(j1, mi1), (mi1, -j1, j1)), (mi, -j1, j1)), JxKet(j2, m2))
    assert qapply(TensorProduct(1, Jz)*TensorProduct(JxKet(j1, m1), JxKet(j2, m2))) == \
        TensorProduct(JxKet(j1, m1), Sum(hbar*mi*WignerD(j2, mi, m2, 0, pi/2, 0)*Sum(WignerD(j2, mi1, mi, 0, pi*Rational(3, 2), 0)*JxKet(j2, mi1), (mi1, -j2, j2)), (mi, -j2, j2)))
    assert qapply(TensorProduct(Jz, 1)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        TensorProduct(Sum(hbar*mi*WignerD(j1, mi, m1, pi*Rational(3, 2), -pi/2, pi/2)*Sum(WignerD(j1, mi1, mi, pi*Rational(3, 2), pi/2, pi/2)*JyKet(j1, mi1), (mi1, -j1, j1)), (mi, -j1, j1)), JyKet(j2, m2))
    assert qapply(TensorProduct(1, Jz)*TensorProduct(JyKet(j1, m1), JyKet(j2, m2))) == \
        TensorProduct(JyKet(j1, m1), Sum(hbar*mi*WignerD(j2, mi, m2, pi*Rational(3, 2), -pi/2, pi/2)*Sum(WignerD(j2, mi1, mi, pi*Rational(3, 2), pi/2, pi/2)*JyKet(j2, mi1), (mi1, -j2, j2)), (mi, -j2, j2)))
    assert qapply(TensorProduct(Jz, 1)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))) == \
        hbar*m1*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))
    assert qapply(TensorProduct(1, Jz)*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))) == \
        hbar*m2*TensorProduct(JzKet(j1, m1), JzKet(j2, m2))


def test_rotation():
    a, b, g = symbols('a b g')
    j, m = symbols('j m')
    #Uncoupled
    answ = [JxKet(1,-1)/2 - sqrt(2)*JxKet(1,0)/2 + JxKet(1,1)/2 ,
       JyKet(1,-1)/2 - sqrt(2)*JyKet(1,0)/2 + JyKet(1,1)/2 ,
       JzKet(1,-1)/2 - sqrt(2)*JzKet(1,0)/2 + JzKet(1,1)/2]
    fun = [state(1, 1) for state in (JxKet, JyKet, JzKet)]
    for state in fun:
        got = qapply(Rotation(0, pi/2, 0)*state)
        assert got in answ
        answ.remove(got)
    assert not answ
    arg = Rotation(a, b, g)*fun[0]
    assert qapply(arg) == (-exp(-I*a)*exp(I*g)*cos(b)*JxKet(1,-1)/2 +
        exp(-I*a)*exp(I*g)*JxKet(1,-1)/2 - sqrt(2)*exp(-I*a)*sin(b)*JxKet(1,0)/2 +
        exp(-I*a)*exp(-I*g)*cos(b)*JxKet(1,1)/2 + exp(-I*a)*exp(-I*g)*JxKet(1,1)/2)
    #dummy effective
    assert str(qapply(Rotation(a, b, g)*JzKet(j, m), dummy=False)) == str(
        qapply(Rotation(a, b, g)*JzKet(j, m), dummy=True)).replace('_','')
    #Coupled
    ans = [JxKetCoupled(1,-1,(1,1))/2 - sqrt(2)*JxKetCoupled(1,0,(1,1))/2 +
       JxKetCoupled(1,1,(1,1))/2 ,
       JyKetCoupled(1,-1,(1,1))/2 - sqrt(2)*JyKetCoupled(1,0,(1,1))/2 +
       JyKetCoupled(1,1,(1,1))/2 ,
       JzKetCoupled(1,-1,(1,1))/2 - sqrt(2)*JzKetCoupled(1,0,(1,1))/2 +
       JzKetCoupled(1,1,(1,1))/2]
    fun = [state(1, 1, (1,1)) for state in (JxKetCoupled, JyKetCoupled, JzKetCoupled)]
    for state in fun:
        got = qapply(Rotation(0, pi/2, 0)*state)
        assert got in ans
        ans.remove(got)
    assert not ans
    arg = Rotation(a, b, g)*fun[0]
    assert qapply(arg) == (
        -exp(-I*a)*exp(I*g)*cos(b)*JxKetCoupled(1,-1,(1,1))/2 +
        exp(-I*a)*exp(I*g)*JxKetCoupled(1,-1,(1,1))/2 -
        sqrt(2)*exp(-I*a)*sin(b)*JxKetCoupled(1,0,(1,1))/2 +
        exp(-I*a)*exp(-I*g)*cos(b)*JxKetCoupled(1,1,(1,1))/2 +
        exp(-I*a)*exp(-I*g)*JxKetCoupled(1,1,(1,1))/2)
    #dummy effective
    assert str(qapply(Rotation(a,b,g)*JzKetCoupled(j,m,(j1,j2)), dummy=False)) == str(
        qapply(Rotation(a,b,g)*JzKetCoupled(j,m,(j1,j2)), dummy=True)).replace('_','')


def test_jzket():
    j, m = symbols('j m')
    # j not integer or half integer
    raises(ValueError, lambda: JzKet(Rational(2, 3), Rational(-1, 3)))
    raises(ValueError, lambda: JzKet(Rational(2, 3), m))
    # j < 0
    raises(ValueError, lambda: JzKet(-1, 1))
    raises(ValueError, lambda: JzKet(-1, m))
    # m not integer or half integer
    raises(ValueError, lambda: JzKet(j, Rational(-1, 3)))
    # abs(m) > j
    raises(ValueError, lambda: JzKet(1, 2))
    raises(ValueError, lambda: JzKet(1, -2))
    # j-m not integer
    raises(ValueError, lambda: JzKet(1, S.Half))


def test_jzketcoupled():
    j, m = symbols('j m')
    # j not integer or half integer
    raises(ValueError, lambda: JzKetCoupled(Rational(2, 3), Rational(-1, 3), (1,)))
    raises(ValueError, lambda: JzKetCoupled(Rational(2, 3), m, (1,)))
    # j < 0
    raises(ValueError, lambda: JzKetCoupled(-1, 1, (1,)))
    raises(ValueError, lambda: JzKetCoupled(-1, m, (1,)))
    # m not integer or half integer
    raises(ValueError, lambda: JzKetCoupled(j, Rational(-1, 3), (1,)))
    # abs(m) > j
    raises(ValueError, lambda: JzKetCoupled(1, 2, (1,)))
    raises(ValueError, lambda: JzKetCoupled(1, -2, (1,)))
    # j-m not integer
    raises(ValueError, lambda: JzKetCoupled(1, S.Half, (1,)))
    # checks types on coupling scheme
    raises(TypeError, lambda: JzKetCoupled(1, 1, 1))
    raises(TypeError, lambda: JzKetCoupled(1, 1, (1,), 1))
    raises(TypeError, lambda: JzKetCoupled(1, 1, (1, 1), (1,)))
    raises(TypeError, lambda: JzKetCoupled(1, 1, (1, 1, 1), (1, 2, 1),
           (1, 3, 1)))
    # checks length of coupling terms
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1,), ((1, 2, 1),)))
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1, 1), ((1, 2),)))
    # all jn are integer or half-integer
    raises(ValueError, lambda: JzKetCoupled(1, 1, (Rational(1, 3), Rational(2, 3))))
    # indices in coupling scheme must be integers
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1, 1), ((S.Half, 1, 2),) ))
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1, 1), ((1, S.Half, 2),) ))
    # indices out of range
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1, 1), ((0, 2, 1),) ))
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1, 1), ((3, 2, 1),) ))
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1, 1), ((1, 0, 1),) ))
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1, 1), ((1, 3, 1),) ))
    # all j values in coupling scheme must by integer or half-integer
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1, 1, 1), ((1, 2, S(
        4)/3), (1, 3, 1)) ))
    # each coupling must satisfy |j1-j2| <= j3 <= j1+j2
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1, 5)))
    raises(ValueError, lambda: JzKetCoupled(5, 1, (1, 1)))
    # final j of coupling must be j of the state
    raises(ValueError, lambda: JzKetCoupled(1, 1, (1, 1), ((1, 2, 2),) ))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_core\_asarray.py ===
"""
Functions in the ``as*array`` family that promote array-likes into arrays.

`require` fits this category despite its name not matching this pattern.
"""
from .multiarray import array, asanyarray
from .overrides import (
    array_function_dispatch,
    finalize_array_function_like,
    set_module,
)

__all__ = ["require"]


POSSIBLE_FLAGS = {
    'C': 'C', 'C_CONTIGUOUS': 'C', 'CONTIGUOUS': 'C',
    'F': 'F', 'F_CONTIGUOUS': 'F', 'FORTRAN': 'F',
    'A': 'A', 'ALIGNED': 'A',
    'W': 'W', 'WRITEABLE': 'W',
    'O': 'O', 'OWNDATA': 'O',
    'E': 'E', 'ENSUREARRAY': 'E'
}


@finalize_array_function_like
@set_module('numpy')
def require(a, dtype=None, requirements=None, *, like=None):
    """
    Return an ndarray of the provided type that satisfies requirements.

    This function is useful to be sure that an array with the correct flags
    is returned for passing to compiled code (perhaps through ctypes).

    Parameters
    ----------
    a : array_like
       The object to be converted to a type-and-requirement-satisfying array.
    dtype : data-type
       The required data-type. If None preserve the current dtype. If your
       application requires the data to be in native byteorder, include
       a byteorder specification as a part of the dtype specification.
    requirements : str or sequence of str
       The requirements list can be any of the following

       * 'F_CONTIGUOUS' ('F') - ensure a Fortran-contiguous array
       * 'C_CONTIGUOUS' ('C') - ensure a C-contiguous array
       * 'ALIGNED' ('A')      - ensure a data-type aligned array
       * 'WRITEABLE' ('W')    - ensure a writable array
       * 'OWNDATA' ('O')      - ensure an array that owns its own data
       * 'ENSUREARRAY', ('E') - ensure a base array, instead of a subclass
    ${ARRAY_FUNCTION_LIKE}

        .. versionadded:: 1.20.0

    Returns
    -------
    out : ndarray
        Array with specified requirements and type if given.

    See Also
    --------
    asarray : Convert input to an ndarray.
    asanyarray : Convert to an ndarray, but pass through ndarray subclasses.
    ascontiguousarray : Convert input to a contiguous array.
    asfortranarray : Convert input to an ndarray with column-major
                     memory order.
    ndarray.flags : Information about the memory layout of the array.

    Notes
    -----
    The returned array will be guaranteed to have the listed requirements
    by making a copy if needed.

    Examples
    --------
    >>> import numpy as np
    >>> x = np.arange(6).reshape(2,3)
    >>> x.flags
      C_CONTIGUOUS : True
      F_CONTIGUOUS : False
      OWNDATA : False
      WRITEABLE : True
      ALIGNED : True
      WRITEBACKIFCOPY : False

    >>> y = np.require(x, dtype=np.float32, requirements=['A', 'O', 'W', 'F'])
    >>> y.flags
      C_CONTIGUOUS : False
      F_CONTIGUOUS : True
      OWNDATA : True
      WRITEABLE : True
      ALIGNED : True
      WRITEBACKIFCOPY : False

    """
    if like is not None:
        return _require_with_like(
            like,
            a,
            dtype=dtype,
            requirements=requirements,
        )

    if not requirements:
        return asanyarray(a, dtype=dtype)

    requirements = {POSSIBLE_FLAGS[x.upper()] for x in requirements}

    if 'E' in requirements:
        requirements.remove('E')
        subok = False
    else:
        subok = True

    order = 'A'
    if requirements >= {'C', 'F'}:
        raise ValueError('Cannot specify both "C" and "F" order')
    elif 'F' in requirements:
        order = 'F'
        requirements.remove('F')
    elif 'C' in requirements:
        order = 'C'
        requirements.remove('C')

    arr = array(a, dtype=dtype, order=order, copy=None, subok=subok)

    for prop in requirements:
        if not arr.flags[prop]:
            return arr.copy(order)
    return arr


_require_with_like = array_function_dispatch()(require)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\requests\help.py ===
"""Module containing bug report helper(s)."""

import json
import platform
import ssl
import sys

import idna
import urllib3

from . import __version__ as requests_version

try:
    import charset_normalizer
except ImportError:
    charset_normalizer = None

try:
    import chardet
except ImportError:
    chardet = None

try:
    from urllib3.contrib import pyopenssl
except ImportError:
    pyopenssl = None
    OpenSSL = None
    cryptography = None
else:
    import cryptography
    import OpenSSL


def _implementation():
    """Return a dict with the Python implementation and version.

    Provide both the name and the version of the Python implementation
    currently running. For example, on CPython 3.10.3 it will return
    {'name': 'CPython', 'version': '3.10.3'}.

    This function works best on CPython and PyPy: in particular, it probably
    doesn't work for Jython or IronPython. Future investigation should be done
    to work out the correct shape of the code for those platforms.
    """
    implementation = platform.python_implementation()

    if implementation == "CPython":
        implementation_version = platform.python_version()
    elif implementation == "PyPy":
        implementation_version = "{}.{}.{}".format(
            sys.pypy_version_info.major,
            sys.pypy_version_info.minor,
            sys.pypy_version_info.micro,
        )
        if sys.pypy_version_info.releaselevel != "final":
            implementation_version = "".join(
                [implementation_version, sys.pypy_version_info.releaselevel]
            )
    elif implementation == "Jython":
        implementation_version = platform.python_version()  # Complete Guess
    elif implementation == "IronPython":
        implementation_version = platform.python_version()  # Complete Guess
    else:
        implementation_version = "Unknown"

    return {"name": implementation, "version": implementation_version}


def info():
    """Generate information for a bug report."""
    try:
        platform_info = {
            "system": platform.system(),
            "release": platform.release(),
        }
    except OSError:
        platform_info = {
            "system": "Unknown",
            "release": "Unknown",
        }

    implementation_info = _implementation()
    urllib3_info = {"version": urllib3.__version__}
    charset_normalizer_info = {"version": None}
    chardet_info = {"version": None}
    if charset_normalizer:
        charset_normalizer_info = {"version": charset_normalizer.__version__}
    if chardet:
        chardet_info = {"version": chardet.__version__}

    pyopenssl_info = {
        "version": None,
        "openssl_version": "",
    }
    if OpenSSL:
        pyopenssl_info = {
            "version": OpenSSL.__version__,
            "openssl_version": f"{OpenSSL.SSL.OPENSSL_VERSION_NUMBER:x}",
        }
    cryptography_info = {
        "version": getattr(cryptography, "__version__", ""),
    }
    idna_info = {
        "version": getattr(idna, "__version__", ""),
    }

    system_ssl = ssl.OPENSSL_VERSION_NUMBER
    system_ssl_info = {"version": f"{system_ssl:x}" if system_ssl is not None else ""}

    return {
        "platform": platform_info,
        "implementation": implementation_info,
        "system_ssl": system_ssl_info,
        "using_pyopenssl": pyopenssl is not None,
        "using_charset_normalizer": chardet is None,
        "pyOpenSSL": pyopenssl_info,
        "urllib3": urllib3_info,
        "chardet": chardet_info,
        "charset_normalizer": charset_normalizer_info,
        "cryptography": cryptography_info,
        "idna": idna_info,
        "requests": {
            "version": requests_version,
        },
    }


def main():
    """Pretty-print the bug information as JSON."""
    print(json.dumps(info(), sort_keys=True, indent=2))


if __name__ == "__main__":
    main()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\remote\command.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.


class Command:
    """Defines constants for the standard WebDriver commands.

    While these constants have no meaning in and of themselves, they are
    used to marshal commands through a service that implements WebDriver's
    remote wire protocol:

        https://w3c.github.io/webdriver/
    """

    NEW_SESSION: str = "newSession"
    DELETE_SESSION: str = "deleteSession"
    NEW_WINDOW: str = "newWindow"
    CLOSE: str = "close"
    QUIT: str = "quit"
    GET: str = "get"
    GO_BACK: str = "goBack"
    GO_FORWARD: str = "goForward"
    REFRESH: str = "refresh"
    ADD_COOKIE: str = "addCookie"
    GET_COOKIE: str = "getCookie"
    GET_ALL_COOKIES: str = "getCookies"
    DELETE_COOKIE: str = "deleteCookie"
    DELETE_ALL_COOKIES: str = "deleteAllCookies"
    FIND_ELEMENT: str = "findElement"
    FIND_ELEMENTS: str = "findElements"
    FIND_CHILD_ELEMENT: str = "findChildElement"
    FIND_CHILD_ELEMENTS: str = "findChildElements"
    CLEAR_ELEMENT: str = "clearElement"
    CLICK_ELEMENT: str = "clickElement"
    SEND_KEYS_TO_ELEMENT: str = "sendKeysToElement"
    W3C_GET_CURRENT_WINDOW_HANDLE: str = "w3cGetCurrentWindowHandle"
    W3C_GET_WINDOW_HANDLES: str = "w3cGetWindowHandles"
    SET_WINDOW_RECT: str = "setWindowRect"
    GET_WINDOW_RECT: str = "getWindowRect"
    SWITCH_TO_WINDOW: str = "switchToWindow"
    SWITCH_TO_FRAME: str = "switchToFrame"
    SWITCH_TO_PARENT_FRAME: str = "switchToParentFrame"
    W3C_GET_ACTIVE_ELEMENT: str = "w3cGetActiveElement"
    GET_CURRENT_URL: str = "getCurrentUrl"
    GET_PAGE_SOURCE: str = "getPageSource"
    GET_TITLE: str = "getTitle"
    W3C_EXECUTE_SCRIPT: str = "w3cExecuteScript"
    W3C_EXECUTE_SCRIPT_ASYNC: str = "w3cExecuteScriptAsync"
    GET_ELEMENT_TEXT: str = "getElementText"
    GET_ELEMENT_TAG_NAME: str = "getElementTagName"
    IS_ELEMENT_SELECTED: str = "isElementSelected"
    IS_ELEMENT_ENABLED: str = "isElementEnabled"
    GET_ELEMENT_RECT: str = "getElementRect"
    GET_ELEMENT_ATTRIBUTE: str = "getElementAttribute"
    GET_ELEMENT_PROPERTY: str = "getElementProperty"
    GET_ELEMENT_VALUE_OF_CSS_PROPERTY: str = "getElementValueOfCssProperty"
    GET_ELEMENT_ARIA_ROLE: str = "getElementAriaRole"
    GET_ELEMENT_ARIA_LABEL: str = "getElementAriaLabel"
    SCREENSHOT: str = "screenshot"
    ELEMENT_SCREENSHOT: str = "elementScreenshot"
    EXECUTE_ASYNC_SCRIPT: str = "executeAsyncScript"
    SET_TIMEOUTS: str = "setTimeouts"
    GET_TIMEOUTS: str = "getTimeouts"
    W3C_MAXIMIZE_WINDOW: str = "w3cMaximizeWindow"
    GET_LOG: str = "getLog"
    GET_AVAILABLE_LOG_TYPES: str = "getAvailableLogTypes"
    FULLSCREEN_WINDOW: str = "fullscreenWindow"
    MINIMIZE_WINDOW: str = "minimizeWindow"
    PRINT_PAGE: str = "printPage"

    # Alerts
    W3C_DISMISS_ALERT: str = "w3cDismissAlert"
    W3C_ACCEPT_ALERT: str = "w3cAcceptAlert"
    W3C_SET_ALERT_VALUE: str = "w3cSetAlertValue"
    W3C_GET_ALERT_TEXT: str = "w3cGetAlertText"

    # Advanced user interactions
    W3C_ACTIONS: str = "actions"
    W3C_CLEAR_ACTIONS: str = "clearActionState"

    # Screen Orientation
    SET_SCREEN_ORIENTATION: str = "setScreenOrientation"
    GET_SCREEN_ORIENTATION: str = "getScreenOrientation"

    # Mobile
    GET_NETWORK_CONNECTION: str = "getNetworkConnection"
    SET_NETWORK_CONNECTION: str = "setNetworkConnection"
    CURRENT_CONTEXT_HANDLE: str = "getCurrentContextHandle"
    CONTEXT_HANDLES: str = "getContextHandles"
    SWITCH_TO_CONTEXT: str = "switchToContext"

    # Web Components
    GET_SHADOW_ROOT: str = "getShadowRoot"
    FIND_ELEMENT_FROM_SHADOW_ROOT: str = "findElementFromShadowRoot"
    FIND_ELEMENTS_FROM_SHADOW_ROOT: str = "findElementsFromShadowRoot"

    # Virtual Authenticator
    ADD_VIRTUAL_AUTHENTICATOR: str = "addVirtualAuthenticator"
    REMOVE_VIRTUAL_AUTHENTICATOR: str = "removeVirtualAuthenticator"
    ADD_CREDENTIAL: str = "addCredential"
    GET_CREDENTIALS: str = "getCredentials"
    REMOVE_CREDENTIAL: str = "removeCredential"
    REMOVE_ALL_CREDENTIALS: str = "removeAllCredentials"
    SET_USER_VERIFIED: str = "setUserVerified"

    # Remote File Management
    UPLOAD_FILE: str = "uploadFile"
    GET_DOWNLOADABLE_FILES: str = "getDownloadableFiles"
    DOWNLOAD_FILE: str = "downloadFile"
    DELETE_DOWNLOADABLE_FILES: str = "deleteDownloadableFiles"

    # Federated Credential Management (FedCM)
    GET_FEDCM_TITLE: str = "getFedcmTitle"
    GET_FEDCM_DIALOG_TYPE: str = "getFedcmDialogType"
    GET_FEDCM_ACCOUNT_LIST: str = "getFedcmAccountList"
    SELECT_FEDCM_ACCOUNT: str = "selectFedcmAccount"
    CLICK_FEDCM_DIALOG_BUTTON: str = "clickFedcmDialogButton"
    CANCEL_FEDCM_DIALOG: str = "cancelFedcmDialog"
    SET_FEDCM_DELAY: str = "setFedcmDelay"
    RESET_FEDCM_COOLDOWN: str = "resetFedcmCooldown"

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\engine\mock.py ===
# engine/mock.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

from __future__ import annotations

from operator import attrgetter
import typing
from typing import Any
from typing import Callable
from typing import cast
from typing import Optional
from typing import Type
from typing import Union

from . import url as _url
from .. import util


if typing.TYPE_CHECKING:
    from .base import Engine
    from .interfaces import _CoreAnyExecuteParams
    from .interfaces import CoreExecuteOptionsParameter
    from .interfaces import Dialect
    from .url import URL
    from ..sql.base import Executable
    from ..sql.ddl import InvokeDDLBase
    from ..sql.schema import HasSchemaAttr
    from ..sql.visitors import Visitable


class MockConnection:
    def __init__(self, dialect: Dialect, execute: Callable[..., Any]):
        self._dialect = dialect
        self._execute_impl = execute

    engine: Engine = cast(Any, property(lambda s: s))
    dialect: Dialect = cast(Any, property(attrgetter("_dialect")))
    name: str = cast(Any, property(lambda s: s._dialect.name))

    def connect(self, **kwargs: Any) -> MockConnection:
        return self

    def schema_for_object(self, obj: HasSchemaAttr) -> Optional[str]:
        return obj.schema

    def execution_options(self, **kw: Any) -> MockConnection:
        return self

    def _run_ddl_visitor(
        self,
        visitorcallable: Type[InvokeDDLBase],
        element: Visitable,
        **kwargs: Any,
    ) -> None:
        kwargs["checkfirst"] = False
        visitorcallable(
            dialect=self.dialect, connection=self, **kwargs
        ).traverse_single(element)

    def execute(
        self,
        obj: Executable,
        parameters: Optional[_CoreAnyExecuteParams] = None,
        execution_options: Optional[CoreExecuteOptionsParameter] = None,
    ) -> Any:
        return self._execute_impl(obj, parameters)


def create_mock_engine(
    url: Union[str, URL], executor: Any, **kw: Any
) -> MockConnection:
    """Create a "mock" engine used for echoing DDL.

    This is a utility function used for debugging or storing the output of DDL
    sequences as generated by :meth:`_schema.MetaData.create_all`
    and related methods.

    The function accepts a URL which is used only to determine the kind of
    dialect to be used, as well as an "executor" callable function which
    will receive a SQL expression object and parameters, which can then be
    echoed or otherwise printed.   The executor's return value is not handled,
    nor does the engine allow regular string statements to be invoked, and
    is therefore only useful for DDL that is sent to the database without
    receiving any results.

    E.g.::

        from sqlalchemy import create_mock_engine


        def dump(sql, *multiparams, **params):
            print(sql.compile(dialect=engine.dialect))


        engine = create_mock_engine("postgresql+psycopg2://", dump)
        metadata.create_all(engine, checkfirst=False)

    :param url: A string URL which typically needs to contain only the
     database backend name.

    :param executor: a callable which receives the arguments ``sql``,
     ``*multiparams`` and ``**params``.  The ``sql`` parameter is typically
     an instance of :class:`.ExecutableDDLElement`, which can then be compiled
     into a string using :meth:`.ExecutableDDLElement.compile`.

    .. versionadded:: 1.4 - the :func:`.create_mock_engine` function replaces
       the previous "mock" engine strategy used with
       :func:`_sa.create_engine`.

    .. seealso::

        :ref:`faq_ddl_as_string`

    """

    # create url.URL object
    u = _url.make_url(url)

    dialect_cls = u.get_dialect()

    dialect_args = {}
    # consume dialect arguments from kwargs
    for k in util.get_cls_kwargs(dialect_cls):
        if k in kw:
            dialect_args[k] = kw.pop(k)

    # create dialect
    dialect = dialect_cls(**dialect_args)

    return MockConnection(dialect, executor)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\benchmarks\bench_symbench.py ===
#!/usr/bin/env python
from sympy.core.random import random
from sympy.core.numbers import (I, Integer, pi)
from sympy.core.symbol import Symbol
from sympy.core.sympify import sympify
from sympy.functions.elementary.miscellaneous import sqrt
from sympy.functions.elementary.trigonometric import sin
from sympy.polys.polytools import factor
from sympy.simplify.simplify import simplify
from sympy.abc import x, y, z
from timeit import default_timer as clock


def bench_R1():
    "real(f(f(f(f(f(f(f(f(f(f(i/2)))))))))))"
    def f(z):
        return sqrt(Integer(1)/3)*z**2 + I/3
    f(f(f(f(f(f(f(f(f(f(I/2)))))))))).as_real_imag()[0]


def bench_R2():
    "Hermite polynomial hermite(15, y)"
    def hermite(n, y):
        if n == 1:
            return 2*y
        if n == 0:
            return 1
        return (2*y*hermite(n - 1, y) - 2*(n - 1)*hermite(n - 2, y)).expand()

    hermite(15, y)


def bench_R3():
    "a = [bool(f==f) for _ in range(10)]"
    f = x + y + z
    [bool(f == f) for _ in range(10)]


def bench_R4():
    # we don't have Tuples
    pass


def bench_R5():
    "blowup(L, 8); L=uniq(L)"
    def blowup(L, n):
        for i in range(n):
            L.append( (L[i] + L[i + 1]) * L[i + 2] )

    def uniq(x):
        v = set(x)
        return v
    L = [x, y, z]
    blowup(L, 8)
    L = uniq(L)


def bench_R6():
    "sum(simplify((x+sin(i))/x+(x-sin(i))/x) for i in range(100))"
    sum(simplify((x + sin(i))/x + (x - sin(i))/x) for i in range(100))


def bench_R7():
    "[f.subs(x, random()) for _ in range(10**4)]"
    f = x**24 + 34*x**12 + 45*x**3 + 9*x**18 + 34*x**10 + 32*x**21
    [f.subs(x, random()) for _ in range(10**4)]


def bench_R8():
    "right(x^2,0,5,10^4)"
    def right(f, a, b, n):
        a = sympify(a)
        b = sympify(b)
        n = sympify(n)
        x = f.atoms(Symbol).pop()
        Deltax = (b - a)/n
        c = a
        est = 0
        for i in range(n):
            c += Deltax
            est += f.subs(x, c)
        return est*Deltax

    right(x**2, 0, 5, 10**4)


def _bench_R9():
    "factor(x^20 - pi^5*y^20)"
    factor(x**20 - pi**5*y**20)


def bench_R10():
    "v = [-pi,-pi+1/10..,pi]"
    def srange(min, max, step):
        v = [min]
        while (max - v[-1]).evalf() > 0:
            v.append(v[-1] + step)
        return v[:-1]
    srange(-pi, pi, sympify(1)/10)


def bench_R11():
    "a = [random() + random()*I for w in [0..1000]]"
    [random() + random()*I for w in range(1000)]


def bench_S1():
    "e=(x+y+z+1)**7;f=e*(e+1);f.expand()"
    e = (x + y + z + 1)**7
    f = e*(e + 1)
    f.expand()


if __name__ == '__main__':
    benchmarks = [
        bench_R1,
        bench_R2,
        bench_R3,
        bench_R5,
        bench_R6,
        bench_R7,
        bench_R8,
        #_bench_R9,
        bench_R10,
        bench_R11,
        #bench_S1,
    ]

    report = []
    for b in benchmarks:
        t = clock()
        b()
        t = clock() - t
        print("%s%65s: %f" % (b.__name__, b.__doc__, t))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\werkzeug\routing\__init__.py ===
"""When it comes to combining multiple controller or view functions
(however you want to call them) you need a dispatcher. A simple way
would be applying regular expression tests on the ``PATH_INFO`` and
calling registered callback functions that return the value then.

This module implements a much more powerful system than simple regular
expression matching because it can also convert values in the URLs and
build URLs.

Here a simple example that creates a URL map for an application with
two subdomains (www and kb) and some URL rules:

.. code-block:: python

    m = Map([
        # Static URLs
        Rule('/', endpoint='static/index'),
        Rule('/about', endpoint='static/about'),
        Rule('/help', endpoint='static/help'),
        # Knowledge Base
        Subdomain('kb', [
            Rule('/', endpoint='kb/index'),
            Rule('/browse/', endpoint='kb/browse'),
            Rule('/browse/<int:id>/', endpoint='kb/browse'),
            Rule('/browse/<int:id>/<int:page>', endpoint='kb/browse')
        ])
    ], default_subdomain='www')

If the application doesn't use subdomains it's perfectly fine to not set
the default subdomain and not use the `Subdomain` rule factory. The
endpoint in the rules can be anything, for example import paths or
unique identifiers. The WSGI application can use those endpoints to get the
handler for that URL.  It doesn't have to be a string at all but it's
recommended.

Now it's possible to create a URL adapter for one of the subdomains and
build URLs:

.. code-block:: python

    c = m.bind('example.com')

    c.build("kb/browse", dict(id=42))
    'http://kb.example.com/browse/42/'

    c.build("kb/browse", dict())
    'http://kb.example.com/browse/'

    c.build("kb/browse", dict(id=42, page=3))
    'http://kb.example.com/browse/42/3'

    c.build("static/about")
    '/about'

    c.build("static/index", force_external=True)
    'http://www.example.com/'

    c = m.bind('example.com', subdomain='kb')

    c.build("static/about")
    'http://www.example.com/about'

The first argument to bind is the server name *without* the subdomain.
Per default it will assume that the script is mounted on the root, but
often that's not the case so you can provide the real mount point as
second argument:

.. code-block:: python

    c = m.bind('example.com', '/applications/example')

The third argument can be the subdomain, if not given the default
subdomain is used.  For more details about binding have a look at the
documentation of the `MapAdapter`.

And here is how you can match URLs:

.. code-block:: python

    c = m.bind('example.com')

    c.match("/")
    ('static/index', {})

    c.match("/about")
    ('static/about', {})

    c = m.bind('example.com', '/', 'kb')

    c.match("/")
    ('kb/index', {})

    c.match("/browse/42/23")
    ('kb/browse', {'id': 42, 'page': 23})

If matching fails you get a ``NotFound`` exception, if the rule thinks
it's a good idea to redirect (for example because the URL was defined
to have a slash at the end but the request was missing that slash) it
will raise a ``RequestRedirect`` exception. Both are subclasses of
``HTTPException`` so you can use those errors as responses in the
application.

If matching succeeded but the URL rule was incompatible to the given
method (for example there were only rules for ``GET`` and ``HEAD`` but
routing tried to match a ``POST`` request) a ``MethodNotAllowed``
exception is raised.
"""

from .converters import AnyConverter as AnyConverter
from .converters import BaseConverter as BaseConverter
from .converters import FloatConverter as FloatConverter
from .converters import IntegerConverter as IntegerConverter
from .converters import PathConverter as PathConverter
from .converters import UnicodeConverter as UnicodeConverter
from .converters import UUIDConverter as UUIDConverter
from .converters import ValidationError as ValidationError
from .exceptions import BuildError as BuildError
from .exceptions import NoMatch as NoMatch
from .exceptions import RequestAliasRedirect as RequestAliasRedirect
from .exceptions import RequestPath as RequestPath
from .exceptions import RequestRedirect as RequestRedirect
from .exceptions import RoutingException as RoutingException
from .exceptions import WebsocketMismatch as WebsocketMismatch
from .map import Map as Map
from .map import MapAdapter as MapAdapter
from .matcher import StateMachineMatcher as StateMachineMatcher
from .rules import EndpointPrefix as EndpointPrefix
from .rules import parse_converter_args as parse_converter_args
from .rules import Rule as Rule
from .rules import RuleFactory as RuleFactory
from .rules import RuleTemplate as RuleTemplate
from .rules import RuleTemplateFactory as RuleTemplateFactory
from .rules import Subdomain as Subdomain
from .rules import Submount as Submount

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\reader\workbook.py ===
# Copyright (c) 2010-2024 openpyxl

from warnings import warn

from openpyxl.xml.functions import fromstring

from openpyxl.packaging.relationship import (
    get_dependents,
    get_rels_path,
    get_rel,
)
from openpyxl.packaging.workbook import WorkbookPackage
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedNameList
from openpyxl.workbook.external_link.external import read_external_link
from openpyxl.pivot.cache import CacheDefinition
from openpyxl.pivot.record import RecordList
from openpyxl.worksheet.print_settings import PrintTitles, PrintArea

from openpyxl.utils.datetime import CALENDAR_MAC_1904


class WorkbookParser:

    _rels = None

    def __init__(self, archive, workbook_part_name, keep_links=True):
        self.archive = archive
        self.workbook_part_name = workbook_part_name
        self.defined_names = DefinedNameList()
        self.wb = Workbook()
        self.keep_links = keep_links
        self.sheets = []


    @property
    def rels(self):
        if self._rels is None:
            self._rels = get_dependents(self.archive, get_rels_path(self.workbook_part_name)).to_dict()
        return self._rels


    def parse(self):
        src = self.archive.read(self.workbook_part_name)
        node = fromstring(src)
        package = WorkbookPackage.from_tree(node)
        if package.properties.date1904:
            self.wb.epoch = CALENDAR_MAC_1904

        self.wb.code_name = package.properties.codeName
        self.wb.active = package.active
        self.wb.views = package.bookViews
        self.sheets = package.sheets
        self.wb.calculation = package.calcPr
        self.caches = package.pivotCaches

        # external links contain cached worksheets and can be very big
        if not self.keep_links:
            package.externalReferences = []

        for ext_ref in package.externalReferences:
            rel = self.rels.get(ext_ref.id)
            self.wb._external_links.append(
                read_external_link(self.archive, rel.Target)
            )

        if package.definedNames:
            self.defined_names = package.definedNames

        self.wb.security = package.workbookProtection


    def find_sheets(self):
        """
        Find all sheets in the workbook and return the link to the source file.

        Older XLSM files sometimes contain invalid sheet elements.
        Warn user when these are removed.
        """

        for sheet in self.sheets:
            if not sheet.id:
                msg = f"File contains an invalid specification for {0}. This will be removed".format(sheet.name)
                warn(msg)
                continue
            yield sheet, self.rels[sheet.id]


    def assign_names(self):
        """
        Bind defined names and other definitions to worksheets or the workbook
        """

        for idx, names in self.defined_names.by_sheet().items():
            if idx == "global":
                self.wb.defined_names = names
                continue

            try:
                sheet = self.wb._sheets[idx]
            except IndexError:
                warn(f"Defined names for sheet index {idx} cannot be located")
                continue

            for name, defn in names.items():
                reserved = defn.is_reserved
                if reserved is None:
                    sheet.defined_names[name] = defn

                elif reserved == "Print_Titles":
                    titles = PrintTitles.from_string(defn.value)
                    sheet._print_rows = titles.rows
                    sheet._print_cols = titles.cols
                elif reserved == "Print_Area":
                    try:
                        sheet._print_area = PrintArea.from_string(defn.value)
                    except TypeError:
                        warn(f"Print area cannot be set to Defined name: {defn.value}.")
                        continue

    @property
    def pivot_caches(self):
        """
        Get PivotCache objects
        """
        d = {}
        for c in self.caches:
            cache = get_rel(self.archive, self.rels, id=c.id, cls=CacheDefinition)
            if cache.deps:
                records = get_rel(self.archive, cache.deps, cache.id, RecordList)
                cache.records = records
            d[c.cacheId] = cache
        return d

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\worksheet\ole.py ===
# Copyright (c) 2010-2024 openpyxl

from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.descriptors import (
    Typed,
    Integer,
    String,
    Set,
    Bool,
    Sequence,
)

from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.xml.constants import SHEET_DRAWING_NS


class ObjectAnchor(Serialisable):

    tagname = "anchor"

    _from = Typed(expected_type=AnchorMarker, namespace=SHEET_DRAWING_NS)
    to = Typed(expected_type=AnchorMarker, namespace=SHEET_DRAWING_NS)
    moveWithCells = Bool(allow_none=True)
    sizeWithCells = Bool(allow_none=True)
    z_order = Integer(allow_none=True, hyphenated=True)


    def __init__(self,
                 _from=None,
                 to=None,
                 moveWithCells=False,
                 sizeWithCells=False,
                 z_order=None,
                ):
        self._from = _from
        self.to = to
        self.moveWithCells = moveWithCells
        self.sizeWithCells = sizeWithCells
        self.z_order = z_order


class ObjectPr(Serialisable):

    tagname = "objectPr"

    anchor = Typed(expected_type=ObjectAnchor, )
    locked = Bool(allow_none=True)
    defaultSize = Bool(allow_none=True)
    _print = Bool(allow_none=True)
    disabled = Bool(allow_none=True)
    uiObject = Bool(allow_none=True)
    autoFill = Bool(allow_none=True)
    autoLine = Bool(allow_none=True)
    autoPict = Bool(allow_none=True)
    macro = String()
    altText = String(allow_none=True)
    dde = Bool(allow_none=True)

    __elements__ = ('anchor',)

    def __init__(self,
                 anchor=None,
                 locked=True,
                 defaultSize=True,
                 _print=True,
                 disabled=False,
                 uiObject=False,
                 autoFill=True,
                 autoLine=True,
                 autoPict=True,
                 macro=None,
                 altText=None,
                 dde=False,
                ):
        self.anchor = anchor
        self.locked = locked
        self.defaultSize = defaultSize
        self._print = _print
        self.disabled = disabled
        self.uiObject = uiObject
        self.autoFill = autoFill
        self.autoLine = autoLine
        self.autoPict = autoPict
        self.macro = macro
        self.altText = altText
        self.dde = dde


class OleObject(Serialisable):

    tagname = "oleObject"

    objectPr = Typed(expected_type=ObjectPr, allow_none=True)
    progId = String(allow_none=True)
    dvAspect = Set(values=(['DVASPECT_CONTENT', 'DVASPECT_ICON']))
    link = String(allow_none=True)
    oleUpdate = Set(values=(['OLEUPDATE_ALWAYS', 'OLEUPDATE_ONCALL']))
    autoLoad = Bool(allow_none=True)
    shapeId = Integer()

    __elements__ = ('objectPr',)

    def __init__(self,
                 objectPr=None,
                 progId=None,
                 dvAspect='DVASPECT_CONTENT',
                 link=None,
                 oleUpdate=None,
                 autoLoad=False,
                 shapeId=None,
                ):
        self.objectPr = objectPr
        self.progId = progId
        self.dvAspect = dvAspect
        self.link = link
        self.oleUpdate = oleUpdate
        self.autoLoad = autoLoad
        self.shapeId = shapeId


class OleObjects(Serialisable):

    tagname = "oleObjects"

    oleObject = Sequence(expected_type=OleObject)

    __elements__ = ('oleObject',)

    def __init__(self,
                 oleObject=(),
                ):
        self.oleObject = oleObject


# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\cli\main_parser.py ===
"""A single place for constructing and exposing the main parser"""

import os
import subprocess
import sys
from typing import List, Optional, Tuple

from pip._internal.build_env import get_runnable_pip
from pip._internal.cli import cmdoptions
from pip._internal.cli.parser import ConfigOptionParser, UpdatingDefaultsHelpFormatter
from pip._internal.commands import commands_dict, get_similar_commands
from pip._internal.exceptions import CommandError
from pip._internal.utils.misc import get_pip_version, get_prog

__all__ = ["create_main_parser", "parse_command"]


def create_main_parser() -> ConfigOptionParser:
    """Creates and returns the main parser for pip's CLI"""

    parser = ConfigOptionParser(
        usage="\n%prog <command> [options]",
        add_help_option=False,
        formatter=UpdatingDefaultsHelpFormatter(),
        name="global",
        prog=get_prog(),
    )
    parser.disable_interspersed_args()

    parser.version = get_pip_version()

    # add the general options
    gen_opts = cmdoptions.make_option_group(cmdoptions.general_group, parser)
    parser.add_option_group(gen_opts)

    # so the help formatter knows
    parser.main = True  # type: ignore

    # create command listing for description
    description = [""] + [
        f"{name:27} {command_info.summary}"
        for name, command_info in commands_dict.items()
    ]
    parser.description = "\n".join(description)

    return parser


def identify_python_interpreter(python: str) -> Optional[str]:
    # If the named file exists, use it.
    # If it's a directory, assume it's a virtual environment and
    # look for the environment's Python executable.
    if os.path.exists(python):
        if os.path.isdir(python):
            # bin/python for Unix, Scripts/python.exe for Windows
            # Try both in case of odd cases like cygwin.
            for exe in ("bin/python", "Scripts/python.exe"):
                py = os.path.join(python, exe)
                if os.path.exists(py):
                    return py
        else:
            return python

    # Could not find the interpreter specified
    return None


def parse_command(args: List[str]) -> Tuple[str, List[str]]:
    parser = create_main_parser()

    # Note: parser calls disable_interspersed_args(), so the result of this
    # call is to split the initial args into the general options before the
    # subcommand and everything else.
    # For example:
    #  args: ['--timeout=5', 'install', '--user', 'INITools']
    #  general_options: ['--timeout==5']
    #  args_else: ['install', '--user', 'INITools']
    general_options, args_else = parser.parse_args(args)

    # --python
    if general_options.python and "_PIP_RUNNING_IN_SUBPROCESS" not in os.environ:
        # Re-invoke pip using the specified Python interpreter
        interpreter = identify_python_interpreter(general_options.python)
        if interpreter is None:
            raise CommandError(
                f"Could not locate Python interpreter {general_options.python}"
            )

        pip_cmd = [
            interpreter,
            get_runnable_pip(),
        ]
        pip_cmd.extend(args)

        # Set a flag so the child doesn't re-invoke itself, causing
        # an infinite loop.
        os.environ["_PIP_RUNNING_IN_SUBPROCESS"] = "1"
        returncode = 0
        try:
            proc = subprocess.run(pip_cmd)
            returncode = proc.returncode
        except (subprocess.SubprocessError, OSError) as exc:
            raise CommandError(f"Failed to run pip under {interpreter}: {exc}")
        sys.exit(returncode)

    # --version
    if general_options.version:
        sys.stdout.write(parser.version)
        sys.stdout.write(os.linesep)
        sys.exit()

    # pip || pip help -> print_help()
    if not args_else or (args_else[0] == "help" and len(args_else) == 1):
        parser.print_help()
        sys.exit()

    # the subcommand name
    cmd_name = args_else[0]

    if cmd_name not in commands_dict:
        guess = get_similar_commands(cmd_name)

        msg = [f'unknown command "{cmd_name}"']
        if guess:
            msg.append(f'maybe you meant "{guess}"')

        raise CommandError(" - ".join(msg))

    # all the args without the subcommand
    cmd_args = args[:]
    cmd_args.remove(cmd_name)

    return cmd_name, cmd_args

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\utils\wheel.py ===
"""Support functions for working with wheel files."""

import logging
from email.message import Message
from email.parser import Parser
from typing import Tuple
from zipfile import BadZipFile, ZipFile

from pip._vendor.packaging.utils import canonicalize_name

from pip._internal.exceptions import UnsupportedWheel

VERSION_COMPATIBLE = (1, 0)


logger = logging.getLogger(__name__)


def parse_wheel(wheel_zip: ZipFile, name: str) -> Tuple[str, Message]:
    """Extract information from the provided wheel, ensuring it meets basic
    standards.

    Returns the name of the .dist-info directory and the parsed WHEEL metadata.
    """
    try:
        info_dir = wheel_dist_info_dir(wheel_zip, name)
        metadata = wheel_metadata(wheel_zip, info_dir)
        version = wheel_version(metadata)
    except UnsupportedWheel as e:
        raise UnsupportedWheel(f"{name} has an invalid wheel, {e}")

    check_compatibility(version, name)

    return info_dir, metadata


def wheel_dist_info_dir(source: ZipFile, name: str) -> str:
    """Returns the name of the contained .dist-info directory.

    Raises AssertionError or UnsupportedWheel if not found, >1 found, or
    it doesn't match the provided name.
    """
    # Zip file path separators must be /
    subdirs = {p.split("/", 1)[0] for p in source.namelist()}

    info_dirs = [s for s in subdirs if s.endswith(".dist-info")]

    if not info_dirs:
        raise UnsupportedWheel(".dist-info directory not found")

    if len(info_dirs) > 1:
        raise UnsupportedWheel(
            "multiple .dist-info directories found: {}".format(", ".join(info_dirs))
        )

    info_dir = info_dirs[0]

    info_dir_name = canonicalize_name(info_dir)
    canonical_name = canonicalize_name(name)
    if not info_dir_name.startswith(canonical_name):
        raise UnsupportedWheel(
            f".dist-info directory {info_dir!r} does not start with {canonical_name!r}"
        )

    return info_dir


def read_wheel_metadata_file(source: ZipFile, path: str) -> bytes:
    try:
        return source.read(path)
        # BadZipFile for general corruption, KeyError for missing entry,
        # and RuntimeError for password-protected files
    except (BadZipFile, KeyError, RuntimeError) as e:
        raise UnsupportedWheel(f"could not read {path!r} file: {e!r}")


def wheel_metadata(source: ZipFile, dist_info_dir: str) -> Message:
    """Return the WHEEL metadata of an extracted wheel, if possible.
    Otherwise, raise UnsupportedWheel.
    """
    path = f"{dist_info_dir}/WHEEL"
    # Zip file path separators must be /
    wheel_contents = read_wheel_metadata_file(source, path)

    try:
        wheel_text = wheel_contents.decode()
    except UnicodeDecodeError as e:
        raise UnsupportedWheel(f"error decoding {path!r}: {e!r}")

    # FeedParser (used by Parser) does not raise any exceptions. The returned
    # message may have .defects populated, but for backwards-compatibility we
    # currently ignore them.
    return Parser().parsestr(wheel_text)


def wheel_version(wheel_data: Message) -> Tuple[int, ...]:
    """Given WHEEL metadata, return the parsed Wheel-Version.
    Otherwise, raise UnsupportedWheel.
    """
    version_text = wheel_data["Wheel-Version"]
    if version_text is None:
        raise UnsupportedWheel("WHEEL is missing Wheel-Version")

    version = version_text.strip()

    try:
        return tuple(map(int, version.split(".")))
    except ValueError:
        raise UnsupportedWheel(f"invalid Wheel-Version: {version!r}")


def check_compatibility(version: Tuple[int, ...], name: str) -> None:
    """Raises errors or warns if called with an incompatible Wheel-Version.

    pip should refuse to install a Wheel-Version that's a major series
    ahead of what it's compatible with (e.g 2.0 > 1.1); and warn when
    installing a version only minor version ahead (e.g 1.2 > 1.1).

    version: a 2-tuple representing a Wheel-Version (Major, Minor)
    name: name of wheel or package to raise exception about

    :raises UnsupportedWheel: when an incompatible Wheel-Version is given
    """
    if version[0] > VERSION_COMPATIBLE[0]:
        raise UnsupportedWheel(
            "{}'s Wheel-Version ({}) is not compatible with this version "
            "of pip".format(name, ".".join(map(str, version)))
        )
    elif version > VERSION_COMPATIBLE:
        logger.warning(
            "Installing from a newer Wheel-Version (%s)",
            ".".join(map(str, version)),
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pycparser\plyparser.py ===
#-----------------------------------------------------------------
# plyparser.py
#
# PLYParser class and other utilities for simplifying programming
# parsers with PLY
#
# Eli Bendersky [https://eli.thegreenplace.net/]
# License: BSD
#-----------------------------------------------------------------

import warnings

class Coord(object):
    """ Coordinates of a syntactic element. Consists of:
            - File name
            - Line number
            - (optional) column number, for the Lexer
    """
    __slots__ = ('file', 'line', 'column', '__weakref__')
    def __init__(self, file, line, column=None):
        self.file = file
        self.line = line
        self.column = column

    def __str__(self):
        str = "%s:%s" % (self.file, self.line)
        if self.column: str += ":%s" % self.column
        return str


class ParseError(Exception): pass


class PLYParser(object):
    def _create_opt_rule(self, rulename):
        """ Given a rule name, creates an optional ply.yacc rule
            for it. The name of the optional rule is
            <rulename>_opt
        """
        optname = rulename + '_opt'

        def optrule(self, p):
            p[0] = p[1]

        optrule.__doc__ = '%s : empty\n| %s' % (optname, rulename)
        optrule.__name__ = 'p_%s' % optname
        setattr(self.__class__, optrule.__name__, optrule)

    def _coord(self, lineno, column=None):
        return Coord(
                file=self.clex.filename,
                line=lineno,
                column=column)

    def _token_coord(self, p, token_idx):
        """ Returns the coordinates for the YaccProduction object 'p' indexed
            with 'token_idx'. The coordinate includes the 'lineno' and
            'column'. Both follow the lex semantic, starting from 1.
        """
        last_cr = p.lexer.lexer.lexdata.rfind('\n', 0, p.lexpos(token_idx))
        if last_cr < 0:
            last_cr = -1
        column = (p.lexpos(token_idx) - (last_cr))
        return self._coord(p.lineno(token_idx), column)

    def _parse_error(self, msg, coord):
        raise ParseError("%s: %s" % (coord, msg))


def parameterized(*params):
    """ Decorator to create parameterized rules.

    Parameterized rule methods must be named starting with 'p_' and contain
    'xxx', and their docstrings may contain 'xxx' and 'yyy'. These will be
    replaced by the given parameter tuples. For example, ``p_xxx_rule()`` with
    docstring 'xxx_rule  : yyy' when decorated with
    ``@parameterized(('id', 'ID'))`` produces ``p_id_rule()`` with the docstring
    'id_rule  : ID'. Using multiple tuples produces multiple rules.
    """
    def decorate(rule_func):
        rule_func._params = params
        return rule_func
    return decorate


def template(cls):
    """ Class decorator to generate rules from parameterized rule templates.

    See `parameterized` for more information on parameterized rules.
    """
    issued_nodoc_warning = False
    for attr_name in dir(cls):
        if attr_name.startswith('p_'):
            method = getattr(cls, attr_name)
            if hasattr(method, '_params'):
                # Remove the template method
                delattr(cls, attr_name)
                # Create parameterized rules from this method; only run this if
                # the method has a docstring. This is to address an issue when
                # pycparser's users are installed in -OO mode which strips
                # docstrings away.
                # See: https://github.com/eliben/pycparser/pull/198/ and
                #      https://github.com/eliben/pycparser/issues/197
                # for discussion.
                if method.__doc__ is not None:
                    _create_param_rules(cls, method)
                elif not issued_nodoc_warning:
                    warnings.warn(
                        'parsing methods must have __doc__ for pycparser to work properly',
                        RuntimeWarning,
                        stacklevel=2)
                    issued_nodoc_warning = True
    return cls


def _create_param_rules(cls, func):
    """ Create ply.yacc rules based on a parameterized rule function

    Generates new methods (one per each pair of parameters) based on the
    template rule function `func`, and attaches them to `cls`. The rule
    function's parameters must be accessible via its `_params` attribute.
    """
    for xxx, yyy in func._params:
        # Use the template method's body for each new method
        def param_rule(self, p):
            func(self, p)

        # Substitute in the params for the grammar rule and function name
        param_rule.__doc__ = func.__doc__.replace('xxx', xxx).replace('yyy', yyy)
        param_rule.__name__ = func.__name__.replace('xxx', xxx)

        # Attach the new method to the class
        setattr(cls, param_rule.__name__, param_rule)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pycparser\ply\ctokens.py ===
# ----------------------------------------------------------------------
# ctokens.py
#
# Token specifications for symbols in ANSI C and C++.  This file is
# meant to be used as a library in other tokenizers.
# ----------------------------------------------------------------------

# Reserved words

tokens = [
    # Literals (identifier, integer constant, float constant, string constant, char const)
    'ID', 'TYPEID', 'INTEGER', 'FLOAT', 'STRING', 'CHARACTER',

    # Operators (+,-,*,/,%,|,&,~,^,<<,>>, ||, &&, !, <, <=, >, >=, ==, !=)
    'PLUS', 'MINUS', 'TIMES', 'DIVIDE', 'MODULO',
    'OR', 'AND', 'NOT', 'XOR', 'LSHIFT', 'RSHIFT',
    'LOR', 'LAND', 'LNOT',
    'LT', 'LE', 'GT', 'GE', 'EQ', 'NE',
    
    # Assignment (=, *=, /=, %=, +=, -=, <<=, >>=, &=, ^=, |=)
    'EQUALS', 'TIMESEQUAL', 'DIVEQUAL', 'MODEQUAL', 'PLUSEQUAL', 'MINUSEQUAL',
    'LSHIFTEQUAL','RSHIFTEQUAL', 'ANDEQUAL', 'XOREQUAL', 'OREQUAL',

    # Increment/decrement (++,--)
    'INCREMENT', 'DECREMENT',

    # Structure dereference (->)
    'ARROW',

    # Ternary operator (?)
    'TERNARY',
    
    # Delimeters ( ) [ ] { } , . ; :
    'LPAREN', 'RPAREN',
    'LBRACKET', 'RBRACKET',
    'LBRACE', 'RBRACE',
    'COMMA', 'PERIOD', 'SEMI', 'COLON',

    # Ellipsis (...)
    'ELLIPSIS',
]
    
# Operators
t_PLUS             = r'\+'
t_MINUS            = r'-'
t_TIMES            = r'\*'
t_DIVIDE           = r'/'
t_MODULO           = r'%'
t_OR               = r'\|'
t_AND              = r'&'
t_NOT              = r'~'
t_XOR              = r'\^'
t_LSHIFT           = r'<<'
t_RSHIFT           = r'>>'
t_LOR              = r'\|\|'
t_LAND             = r'&&'
t_LNOT             = r'!'
t_LT               = r'<'
t_GT               = r'>'
t_LE               = r'<='
t_GE               = r'>='
t_EQ               = r'=='
t_NE               = r'!='

# Assignment operators

t_EQUALS           = r'='
t_TIMESEQUAL       = r'\*='
t_DIVEQUAL         = r'/='
t_MODEQUAL         = r'%='
t_PLUSEQUAL        = r'\+='
t_MINUSEQUAL       = r'-='
t_LSHIFTEQUAL      = r'<<='
t_RSHIFTEQUAL      = r'>>='
t_ANDEQUAL         = r'&='
t_OREQUAL          = r'\|='
t_XOREQUAL         = r'\^='

# Increment/decrement
t_INCREMENT        = r'\+\+'
t_DECREMENT        = r'--'

# ->
t_ARROW            = r'->'

# ?
t_TERNARY          = r'\?'

# Delimeters
t_LPAREN           = r'\('
t_RPAREN           = r'\)'
t_LBRACKET         = r'\['
t_RBRACKET         = r'\]'
t_LBRACE           = r'\{'
t_RBRACE           = r'\}'
t_COMMA            = r','
t_PERIOD           = r'\.'
t_SEMI             = r';'
t_COLON            = r':'
t_ELLIPSIS         = r'\.\.\.'

# Identifiers
t_ID = r'[A-Za-z_][A-Za-z0-9_]*'

# Integer literal
t_INTEGER = r'\d+([uU]|[lL]|[uU][lL]|[lL][uU])?'

# Floating literal
t_FLOAT = r'((\d+)(\.\d+)(e(\+|-)?(\d+))? | (\d+)e(\+|-)?(\d+))([lL]|[fF])?'

# String literal
t_STRING = r'\"([^\\\n]|(\\.))*?\"'

# Character constant 'c' or L'c'
t_CHARACTER = r'(L)?\'([^\\\n]|(\\.))*?\''

# Comment (C-Style)
def t_COMMENT(t):
    r'/\*(.|\n)*?\*/'
    t.lexer.lineno += t.value.count('\n')
    return t

# Comment (C++-Style)
def t_CPPCOMMENT(t):
    r'//.*\n'
    t.lexer.lineno += 1
    return t


    




# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pytz\tzfile.py ===
'''
$Id: tzfile.py,v 1.8 2004/06/03 00:15:24 zenzen Exp $
'''

from datetime import datetime
from struct import unpack, calcsize

from pytz.tzinfo import StaticTzInfo, DstTzInfo, memorized_ttinfo
from pytz.tzinfo import memorized_datetime, memorized_timedelta


def _byte_string(s):
    """Cast a string or byte string to an ASCII byte string."""
    return s.encode('ASCII')

_NULL = _byte_string('\0')


def _std_string(s):
    """Cast a string or byte string to an ASCII string."""
    return str(s.decode('ASCII'))


def build_tzinfo(zone, fp):
    head_fmt = '>4s c 15x 6l'
    head_size = calcsize(head_fmt)
    (magic, format, ttisgmtcnt, ttisstdcnt, leapcnt, timecnt,
        typecnt, charcnt) = unpack(head_fmt, fp.read(head_size))

    # Make sure it is a tzfile(5) file
    assert magic == _byte_string('TZif'), 'Got magic %s' % repr(magic)

    # Read out the transition times, localtime indices and ttinfo structures.
    data_fmt = '>%(timecnt)dl %(timecnt)dB %(ttinfo)s %(charcnt)ds' % dict(
        timecnt=timecnt, ttinfo='lBB' * typecnt, charcnt=charcnt)
    data_size = calcsize(data_fmt)
    data = unpack(data_fmt, fp.read(data_size))

    # make sure we unpacked the right number of values
    assert len(data) == 2 * timecnt + 3 * typecnt + 1
    transitions = [memorized_datetime(trans)
                   for trans in data[:timecnt]]
    lindexes = list(data[timecnt:2 * timecnt])
    ttinfo_raw = data[2 * timecnt:-1]
    tznames_raw = data[-1]
    del data

    # Process ttinfo into separate structs
    ttinfo = []
    tznames = {}
    i = 0
    while i < len(ttinfo_raw):
        # have we looked up this timezone name yet?
        tzname_offset = ttinfo_raw[i + 2]
        if tzname_offset not in tznames:
            nul = tznames_raw.find(_NULL, tzname_offset)
            if nul < 0:
                nul = len(tznames_raw)
            tznames[tzname_offset] = _std_string(
                tznames_raw[tzname_offset:nul])
        ttinfo.append((ttinfo_raw[i],
                       bool(ttinfo_raw[i + 1]),
                       tznames[tzname_offset]))
        i += 3

    # Now build the timezone object
    if len(ttinfo) == 1 or len(transitions) == 0:
        ttinfo[0][0], ttinfo[0][2]
        cls = type(zone, (StaticTzInfo,), dict(
            zone=zone,
            _utcoffset=memorized_timedelta(ttinfo[0][0]),
            _tzname=ttinfo[0][2]))
    else:
        # Early dates use the first standard time ttinfo
        i = 0
        while ttinfo[i][1]:
            i += 1
        if ttinfo[i] == ttinfo[lindexes[0]]:
            transitions[0] = datetime.min
        else:
            transitions.insert(0, datetime.min)
            lindexes.insert(0, i)

        # calculate transition info
        transition_info = []
        for i in range(len(transitions)):
            inf = ttinfo[lindexes[i]]
            utcoffset = inf[0]
            if not inf[1]:
                dst = 0
            else:
                for j in range(i - 1, -1, -1):
                    prev_inf = ttinfo[lindexes[j]]
                    if not prev_inf[1]:
                        break
                dst = inf[0] - prev_inf[0]  # dst offset

                # Bad dst? Look further. DST > 24 hours happens when
                # a timzone has moved across the international dateline.
                if dst <= 0 or dst > 3600 * 3:
                    for j in range(i + 1, len(transitions)):
                        stdinf = ttinfo[lindexes[j]]
                        if not stdinf[1]:
                            dst = inf[0] - stdinf[0]
                            if dst > 0:
                                break  # Found a useful std time.

            tzname = inf[2]

            # Round utcoffset and dst to the nearest minute or the
            # datetime library will complain. Conversions to these timezones
            # might be up to plus or minus 30 seconds out, but it is
            # the best we can do.
            utcoffset = int((utcoffset + 30) // 60) * 60
            dst = int((dst + 30) // 60) * 60
            transition_info.append(memorized_ttinfo(utcoffset, dst, tzname))

        cls = type(zone, (DstTzInfo,), dict(
            zone=zone,
            _utc_transition_times=transitions,
            _transition_info=transition_info))

    return cls()

if __name__ == '__main__':
    import os.path
    from pprint import pprint
    base = os.path.join(os.path.dirname(__file__), 'zoneinfo')
    tz = build_tzinfo('Australia/Melbourne',
                      open(os.path.join(base, 'Australia', 'Melbourne'), 'rb'))
    tz = build_tzinfo('US/Eastern',
                      open(os.path.join(base, 'US', 'Eastern'), 'rb'))
    pprint(tz._utc_transition_times)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\firefox\options.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.
from typing import Any, Dict, Optional, Union

from typing_extensions import deprecated

from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.options import ArgOptions
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile


class Log:
    def __init__(self) -> None:
        self.level = None

    def to_capabilities(self) -> dict:
        if self.level:
            return {"log": {"level": self.level}}
        return {}


class Options(ArgOptions):
    KEY = "moz:firefoxOptions"

    def __init__(self) -> None:
        super().__init__()
        self._binary_location = ""
        self._preferences: dict = {}
        # https://fxdx.dev/deprecating-cdp-support-in-firefox-embracing-the-future-with-webdriver-bidi/.
        # Enable BiDi only
        self._preferences["remote.active-protocols"] = 1
        self._profile: Optional[FirefoxProfile] = None
        self.log = Log()

    @property
    @deprecated("use binary_location instead")
    def binary(self) -> FirefoxBinary:
        """Returns the FirefoxBinary instance."""
        return FirefoxBinary(self._binary_location)

    @binary.setter
    @deprecated("use binary_location instead")
    def binary(self, new_binary: Union[str, FirefoxBinary]) -> None:
        """Sets location of the browser binary, either by string or
        ``FirefoxBinary`` instance."""
        if isinstance(new_binary, FirefoxBinary):
            new_binary = new_binary._start_cmd
        self.binary_location = str(new_binary)

    @property
    def binary_location(self) -> str:
        """:Returns: The location of the binary."""
        return self._binary_location

    @binary_location.setter  # noqa
    def binary_location(self, value: str) -> None:
        """Sets the location of the browser binary by string."""
        if not isinstance(value, str):
            raise TypeError(self.BINARY_LOCATION_ERROR)
        self._binary_location = value

    @property
    def preferences(self) -> dict:
        """:Returns: A dict of preferences."""
        return self._preferences

    def set_preference(self, name: str, value: Union[str, int, bool]):
        """Sets a preference."""
        self._preferences[name] = value

    @property
    def profile(self) -> Optional[FirefoxProfile]:
        """:Returns: The Firefox profile to use."""
        return self._profile

    @profile.setter
    def profile(self, new_profile: Union[str, FirefoxProfile]) -> None:
        """Sets location of the browser profile to use, either by string or
        ``FirefoxProfile``."""
        if not isinstance(new_profile, FirefoxProfile):
            new_profile = FirefoxProfile(new_profile)
        self._profile = new_profile

    def enable_mobile(
        self, android_package: Optional[str] = "org.mozilla.firefox", android_activity=None, device_serial=None
    ):
        super().enable_mobile(android_package, android_activity, device_serial)

    def to_capabilities(self) -> dict:
        """Marshals the Firefox options to a `moz:firefoxOptions` object."""
        # This intentionally looks at the internal properties
        # so if a binary or profile has _not_ been set,
        # it will defer to geckodriver to find the system Firefox
        # and generate a fresh profile.
        caps = self._caps
        opts: Dict[str, Any] = {}

        if self._binary_location:
            opts["binary"] = self._binary_location
        if self._preferences:
            opts["prefs"] = self._preferences
        if self._profile:
            opts["profile"] = self._profile.encoded
        if self._arguments:
            opts["args"] = self._arguments
        if self.mobile_options:
            opts.update(self.mobile_options)

        opts.update(self.log.to_capabilities())

        if opts:
            caps[Options.KEY] = opts

        return caps

    @property
    def default_capabilities(self) -> dict:
        return DesiredCapabilities.FIREFOX.copy()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\__init__.py ===
"""Trio - A friendly Python library for async concurrency and I/O"""

from __future__ import annotations

from typing import TYPE_CHECKING

# General layout:
#
# trio/_core/... is the self-contained core library. It does various
# shenanigans to export a consistent "core API", but parts of the core API are
# too low-level to be recommended for regular use.
#
# trio/*.py define a set of more usable tools on top of this. They import from
# trio._core and from each other.
#
# This file pulls together the friendly public API, by re-exporting the more
# innocuous bits of the _core API + the higher-level tools from trio/*.py.
#
# Uses `from x import y as y` for compatibility with `pyright --verifytypes` (#2625)
#
# must be imported early to avoid circular import
from ._core import TASK_STATUS_IGNORED as TASK_STATUS_IGNORED  # isort: split

# Submodules imported by default
from . import abc, from_thread, lowlevel, socket, to_thread
from ._channel import (
    MemoryChannelStatistics as MemoryChannelStatistics,
    MemoryReceiveChannel as MemoryReceiveChannel,
    MemorySendChannel as MemorySendChannel,
    as_safe_channel as as_safe_channel,
    open_memory_channel as open_memory_channel,
)
from ._core import (
    BrokenResourceError as BrokenResourceError,
    BusyResourceError as BusyResourceError,
    Cancelled as Cancelled,
    CancelScope as CancelScope,
    ClosedResourceError as ClosedResourceError,
    EndOfChannel as EndOfChannel,
    Nursery as Nursery,
    RunFinishedError as RunFinishedError,
    TaskStatus as TaskStatus,
    TrioInternalError as TrioInternalError,
    WouldBlock as WouldBlock,
    current_effective_deadline as current_effective_deadline,
    current_time as current_time,
    open_nursery as open_nursery,
    run as run,
)
from ._deprecate import TrioDeprecationWarning as TrioDeprecationWarning
from ._dtls import (
    DTLSChannel as DTLSChannel,
    DTLSChannelStatistics as DTLSChannelStatistics,
    DTLSEndpoint as DTLSEndpoint,
)
from ._file_io import open_file as open_file, wrap_file as wrap_file
from ._highlevel_generic import (
    StapledStream as StapledStream,
    aclose_forcefully as aclose_forcefully,
)
from ._highlevel_open_tcp_listeners import (
    open_tcp_listeners as open_tcp_listeners,
    serve_tcp as serve_tcp,
)
from ._highlevel_open_tcp_stream import open_tcp_stream as open_tcp_stream
from ._highlevel_open_unix_stream import open_unix_socket as open_unix_socket
from ._highlevel_serve_listeners import serve_listeners as serve_listeners
from ._highlevel_socket import (
    SocketListener as SocketListener,
    SocketStream as SocketStream,
)
from ._highlevel_ssl_helpers import (
    open_ssl_over_tcp_listeners as open_ssl_over_tcp_listeners,
    open_ssl_over_tcp_stream as open_ssl_over_tcp_stream,
    serve_ssl_over_tcp as serve_ssl_over_tcp,
)
from ._path import Path as Path, PosixPath as PosixPath, WindowsPath as WindowsPath
from ._signals import open_signal_receiver as open_signal_receiver
from ._ssl import (
    NeedHandshakeError as NeedHandshakeError,
    SSLListener as SSLListener,
    SSLStream as SSLStream,
)
from ._subprocess import Process as Process, run_process as run_process
from ._sync import (
    CapacityLimiter as CapacityLimiter,
    CapacityLimiterStatistics as CapacityLimiterStatistics,
    Condition as Condition,
    ConditionStatistics as ConditionStatistics,
    Event as Event,
    EventStatistics as EventStatistics,
    Lock as Lock,
    LockStatistics as LockStatistics,
    Semaphore as Semaphore,
    StrictFIFOLock as StrictFIFOLock,
)
from ._timeouts import (
    TooSlowError as TooSlowError,
    fail_after as fail_after,
    fail_at as fail_at,
    move_on_after as move_on_after,
    move_on_at as move_on_at,
    sleep as sleep,
    sleep_forever as sleep_forever,
    sleep_until as sleep_until,
)
from ._version import __version__ as __version__

# Not imported by default, but mentioned here so static analysis tools like
# pylint will know that it exists.
if TYPE_CHECKING:
    from . import testing

from . import _deprecate as _deprecate

_deprecate.deprecate_attributes(__name__, {})

# Having the public path in .__module__ attributes is important for:
# - exception names in printed tracebacks
# - sphinx :show-inheritance:
# - deprecation warnings
# - pickle
# - probably other stuff
from ._util import fixup_module_metadata

fixup_module_metadata(__name__, globals())
fixup_module_metadata(lowlevel.__name__, lowlevel.__dict__)
fixup_module_metadata(socket.__name__, socket.__dict__)
fixup_module_metadata(abc.__name__, abc.__dict__)
fixup_module_metadata(from_thread.__name__, from_thread.__dict__)
fixup_module_metadata(to_thread.__name__, to_thread.__dict__)
del fixup_module_metadata
del TYPE_CHECKING

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\h11\_abnf.py ===
# We use native strings for all the re patterns, to take advantage of string
# formatting, and then convert to bytestrings when compiling the final re
# objects.

# https://svn.tools.ietf.org/svn/wg/httpbis/specs/rfc7230.html#whitespace
#  OWS            = *( SP / HTAB )
#                 ; optional whitespace
OWS = r"[ \t]*"

# https://svn.tools.ietf.org/svn/wg/httpbis/specs/rfc7230.html#rule.token.separators
#   token          = 1*tchar
#
#   tchar          = "!" / "#" / "$" / "%" / "&" / "'" / "*"
#                  / "+" / "-" / "." / "^" / "_" / "`" / "|" / "~"
#                  / DIGIT / ALPHA
#                  ; any VCHAR, except delimiters
token = r"[-!#$%&'*+.^_`|~0-9a-zA-Z]+"

# https://svn.tools.ietf.org/svn/wg/httpbis/specs/rfc7230.html#header.fields
#  field-name     = token
field_name = token

# The standard says:
#
#  field-value    = *( field-content / obs-fold )
#  field-content  = field-vchar [ 1*( SP / HTAB ) field-vchar ]
#  field-vchar    = VCHAR / obs-text
#  obs-fold       = CRLF 1*( SP / HTAB )
#                 ; obsolete line folding
#                 ; see Section 3.2.4
#
# https://tools.ietf.org/html/rfc5234#appendix-B.1
#
#   VCHAR          =  %x21-7E
#                  ; visible (printing) characters
#
# https://svn.tools.ietf.org/svn/wg/httpbis/specs/rfc7230.html#rule.quoted-string
#   obs-text       = %x80-FF
#
# However, the standard definition of field-content is WRONG! It disallows
# fields containing a single visible character surrounded by whitespace,
# e.g. "foo a bar".
#
# See: https://www.rfc-editor.org/errata_search.php?rfc=7230&eid=4189
#
# So our definition of field_content attempts to fix it up...
#
# Also, we allow lots of control characters, because apparently people assume
# that they're legal in practice (e.g., google analytics makes cookies with
# \x01 in them!):
#   https://github.com/python-hyper/h11/issues/57
# We still don't allow NUL or whitespace, because those are often treated as
# meta-characters and letting them through can lead to nasty issues like SSRF.
vchar = r"[\x21-\x7e]"
vchar_or_obs_text = r"[^\x00\s]"
field_vchar = vchar_or_obs_text
field_content = r"{field_vchar}+(?:[ \t]+{field_vchar}+)*".format(**globals())

# We handle obs-fold at a different level, and our fixed-up field_content
# already grows to swallow the whole value, so ? instead of *
field_value = r"({field_content})?".format(**globals())

#  header-field   = field-name ":" OWS field-value OWS
header_field = (
    r"(?P<field_name>{field_name})"
    r":"
    r"{OWS}"
    r"(?P<field_value>{field_value})"
    r"{OWS}".format(**globals())
)

# https://svn.tools.ietf.org/svn/wg/httpbis/specs/rfc7230.html#request.line
#
#   request-line   = method SP request-target SP HTTP-version CRLF
#   method         = token
#   HTTP-version   = HTTP-name "/" DIGIT "." DIGIT
#   HTTP-name      = %x48.54.54.50 ; "HTTP", case-sensitive
#
# request-target is complicated (see RFC 7230 sec 5.3) -- could be path, full
# URL, host+port (for connect), or even "*", but in any case we are guaranteed
# that it contists of the visible printing characters.
method = token
request_target = r"{vchar}+".format(**globals())
http_version = r"HTTP/(?P<http_version>[0-9]\.[0-9])"
request_line = (
    r"(?P<method>{method})"
    r" "
    r"(?P<target>{request_target})"
    r" "
    r"{http_version}".format(**globals())
)

# https://svn.tools.ietf.org/svn/wg/httpbis/specs/rfc7230.html#status.line
#
#   status-line = HTTP-version SP status-code SP reason-phrase CRLF
#   status-code    = 3DIGIT
#   reason-phrase  = *( HTAB / SP / VCHAR / obs-text )
status_code = r"[0-9]{3}"
reason_phrase = r"([ \t]|{vchar_or_obs_text})*".format(**globals())
status_line = (
    r"{http_version}"
    r" "
    r"(?P<status_code>{status_code})"
    # However, there are apparently a few too many servers out there that just
    # leave out the reason phrase:
    #   https://github.com/scrapy/scrapy/issues/345#issuecomment-281756036
    #   https://github.com/seanmonstar/httparse/issues/29
    # so make it optional. ?: is a non-capturing group.
    r"(?: (?P<reason>{reason_phrase}))?".format(**globals())
)

HEXDIG = r"[0-9A-Fa-f]"
# Actually
#
#      chunk-size     = 1*HEXDIG
#
# but we impose an upper-limit to avoid ridiculosity. len(str(2**64)) == 20
chunk_size = r"({HEXDIG}){{1,20}}".format(**globals())
# Actually
#
#     chunk-ext      = *( ";" chunk-ext-name [ "=" chunk-ext-val ] )
#
# but we aren't parsing the things so we don't really care.
chunk_ext = r";.*"
chunk_header = (
    r"(?P<chunk_size>{chunk_size})"
    r"(?P<chunk_ext>{chunk_ext})?"
    r"{OWS}\r\n".format(
        **globals()
    )  # Even though the specification does not allow for extra whitespaces,
    # we are lenient with trailing whitespaces because some servers on the wild use it.
)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\quantization\execution_providers\qnn\fusion_lpnorm.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation. All rights reserved.
# Licensed under the MIT License. See License.txt in the project root for
# license information.
# --------------------------------------------------------------------------
from __future__ import annotations

import onnx

from ...fusions import Fusion
from ...onnx_model import ONNXModel


class FusionLpNormalization(Fusion):
    def __init__(self, model: ONNXModel, epsilon: float = 1e-12):
        super().__init__(model, "LpNormalization", "ReduceL2")
        self.epsilon = epsilon

    def fuse(
        self,
        reduce_node: onnx.NodeProto,
        input_name_to_nodes: dict[str, list[onnx.NodeProto]],
        output_name_to_node: dict[str, onnx.NodeProto],
    ):
        """
        Interface function that tries to fuse a node sequence containing a ReduceL2 node into a single
        LpNormalization node.

        Pattern 1:
                    [root] --> ReduceL2 -----> Clip  --> Expand ----> Div -->
                       |      (axis=-1)    (min=epsilon) (shape=root)  ^
                       |   (keepdims=True)                             |
                       |                                               |
                       +-----------------------------------------------+
        Notes:
          - ReduceL2 must use the last axis, and keepdims == True
          - Clip must only have a min attribute that is ~1e-12
          - Expand must restore the shape to root.shape
          - The output of Expand must be the second input to Div.
        """
        if reduce_node.output[0] not in input_name_to_nodes:
            return

        # ReduceL2 must have one Clip child
        children = input_name_to_nodes[reduce_node.output[0]]
        if len(children) != 1 or children[0].op_type != "Clip":
            return

        # ReduceL2 must have keepdims == True
        keepdims = self.get_node_attribute(reduce_node, "keepdims")
        if not keepdims:
            return

        # ReduceL2 axes must refer only to the last dimension.
        # Axes became an input in opset 18. Before then, axes was an attribute
        reduce_input_ttype = self.model.get_tensor_type(reduce_node.input[0])
        if not reduce_input_ttype:
            return

        reduce_input_shape = self.tensor_shape_to_list(reduce_input_ttype)
        if not reduce_input_shape:
            return

        axes = self.get_node_attribute(reduce_node, "axes")
        if not axes and len(reduce_node.input) > 1:
            axes = self.model.get_constant_value(reduce_node.input[1])

        if not axes or len(axes) != 1:
            return

        last_dim = len(reduce_input_shape) - 1
        if axes[0] != -1 and axes[0] != last_dim:
            return

        # Clip node must have a min attribute approximately equal to 1e-12
        clip_node = children[0]
        clip_min = self.get_node_attribute(clip_node, "min")
        if clip_min is None and len(clip_node.input) > 1:
            clip_min = self.model.get_constant_value(clip_node.input[1])

        clip_max = self.get_node_attribute(clip_node, "max")  # TODO: clip_max could be FLOAT_MAX
        if clip_max is None and len(clip_node.input) > 2:
            clip_max = self.model.get_constant_value(clip_node.input[2])

        if not (clip_max is None and clip_min is not None and clip_min > 0 and abs(clip_min - self.epsilon) < 1e-13):
            return

        if clip_node.output[0] not in input_name_to_nodes:
            return

        # Clip must have a single Expand child.
        children = input_name_to_nodes[clip_node.output[0]]
        if len(children) != 1 or children[0].op_type != "Expand":
            return

        expand_node = children[0]
        if expand_node.output[0] not in input_name_to_nodes:
            return

        # Expand must have a single Div child
        children = input_name_to_nodes[expand_node.output[0]]
        if len(children) != 1 or children[0].op_type != "Div":
            return

        div_node = children[0]

        # The first input to Div must be the root of the subgraph (i.e., reduce_node.input[0])
        # The second input to Div must be the output of the Expand.
        # As long as these two inputs go to the same Div node, then ONNX validation will ensure that
        # their shapes match.
        if div_node.input[0] != reduce_node.input[0]:
            return
        if div_node.input[1] != expand_node.output[0]:
            return

        subgraph_input = reduce_node.input[0]
        subgraph_output = div_node.output[0]

        subgraph_nodes = [reduce_node, clip_node, expand_node, div_node]
        if not self.is_safe_to_fuse_nodes(subgraph_nodes, [subgraph_output], input_name_to_nodes, output_name_to_node):
            return

        self.nodes_to_remove.extend(subgraph_nodes)
        fused_node = onnx.helper.make_node(
            self.fused_op_type,
            name=self.create_unique_node_name(),
            inputs=[subgraph_input],
            outputs=[subgraph_output],
            p=2,
            axis=-1,
        )
        self.nodes_to_add.append(fused_node)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\sql\_dml_constructors.py ===
# sql/_dml_constructors.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

from __future__ import annotations

from typing import TYPE_CHECKING

from .dml import Delete
from .dml import Insert
from .dml import Update

if TYPE_CHECKING:
    from ._typing import _DMLTableArgument


def insert(table: _DMLTableArgument) -> Insert:
    """Construct an :class:`_expression.Insert` object.

    E.g.::

        from sqlalchemy import insert

        stmt = insert(user_table).values(name="username", fullname="Full Username")

    Similar functionality is available via the
    :meth:`_expression.TableClause.insert` method on
    :class:`_schema.Table`.

    .. seealso::

        :ref:`tutorial_core_insert` - in the :ref:`unified_tutorial`


    :param table: :class:`_expression.TableClause`
     which is the subject of the
     insert.

    :param values: collection of values to be inserted; see
     :meth:`_expression.Insert.values`
     for a description of allowed formats here.
     Can be omitted entirely; a :class:`_expression.Insert` construct
     will also dynamically render the VALUES clause at execution time
     based on the parameters passed to :meth:`_engine.Connection.execute`.

    :param inline: if True, no attempt will be made to retrieve the
     SQL-generated default values to be provided within the statement;
     in particular,
     this allows SQL expressions to be rendered 'inline' within the
     statement without the need to pre-execute them beforehand; for
     backends that support "returning", this turns off the "implicit
     returning" feature for the statement.

    If both :paramref:`_expression.insert.values` and compile-time bind
    parameters are present, the compile-time bind parameters override the
    information specified within :paramref:`_expression.insert.values` on a
    per-key basis.

    The keys within :paramref:`_expression.Insert.values` can be either
    :class:`~sqlalchemy.schema.Column` objects or their string
    identifiers. Each key may reference one of:

    * a literal data value (i.e. string, number, etc.);
    * a Column object;
    * a SELECT statement.

    If a ``SELECT`` statement is specified which references this
    ``INSERT`` statement's table, the statement will be correlated
    against the ``INSERT`` statement.

    .. seealso::

        :ref:`tutorial_core_insert` - in the :ref:`unified_tutorial`

    """  # noqa: E501
    return Insert(table)


def update(table: _DMLTableArgument) -> Update:
    r"""Construct an :class:`_expression.Update` object.

    E.g.::

        from sqlalchemy import update

        stmt = (
            update(user_table).where(user_table.c.id == 5).values(name="user #5")
        )

    Similar functionality is available via the
    :meth:`_expression.TableClause.update` method on
    :class:`_schema.Table`.

    :param table: A :class:`_schema.Table`
     object representing the database
     table to be updated.


    .. seealso::

        :ref:`tutorial_core_update_delete` - in the :ref:`unified_tutorial`


    """  # noqa: E501
    return Update(table)


def delete(table: _DMLTableArgument) -> Delete:
    r"""Construct :class:`_expression.Delete` object.

    E.g.::

        from sqlalchemy import delete

        stmt = delete(user_table).where(user_table.c.id == 5)

    Similar functionality is available via the
    :meth:`_expression.TableClause.delete` method on
    :class:`_schema.Table`.

    :param table: The table to delete rows from.

    .. seealso::

        :ref:`tutorial_core_update_delete` - in the :ref:`unified_tutorial`


    """
    return Delete(table)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\wtforms\meta.py ===
from wtforms import i18n
from wtforms.utils import WebobInputWrapper
from wtforms.widgets.core import clean_key


class DefaultMeta:
    """
    This is the default Meta class which defines all the default values and
    therefore also the 'API' of the class Meta interface.
    """

    # -- Basic form primitives

    def bind_field(self, form, unbound_field, options):
        """
        bind_field allows potential customization of how fields are bound.

        The default implementation simply passes the options to
        :meth:`UnboundField.bind`.

        :param form: The form.
        :param unbound_field: The unbound field.
        :param options:
            A dictionary of options which are typically passed to the field.

        :return: A bound field
        """
        return unbound_field.bind(form=form, **options)

    def wrap_formdata(self, form, formdata):
        """
        wrap_formdata allows doing custom wrappers of WTForms formdata.

        The default implementation detects webob-style multidicts and wraps
        them, otherwise passes formdata back un-changed.

        :param form: The form.
        :param formdata: Form data.
        :return: A form-input wrapper compatible with WTForms.
        """
        if formdata is not None and not hasattr(formdata, "getlist"):
            if hasattr(formdata, "getall"):
                return WebobInputWrapper(formdata)
            else:
                raise TypeError(
                    "formdata should be a multidict-type wrapper that"
                    " supports the 'getlist' method"
                )
        return formdata

    def render_field(self, field, render_kw):
        """
        render_field allows customization of how widget rendering is done.

        The default implementation calls ``field.widget(field, **render_kw)``
        """

        render_kw = {clean_key(k): v for k, v in render_kw.items()}

        other_kw = getattr(field, "render_kw", None)
        if other_kw is not None:
            other_kw = {clean_key(k): v for k, v in other_kw.items()}
            render_kw = dict(other_kw, **render_kw)
        return field.widget(field, **render_kw)

    # -- CSRF

    csrf = False
    csrf_field_name = "csrf_token"
    csrf_secret = None
    csrf_context = None
    csrf_class = None

    def build_csrf(self, form):
        """
        Build a CSRF implementation. This is called once per form instance.

        The default implementation builds the class referenced to by
        :attr:`csrf_class` with zero arguments. If `csrf_class` is ``None``,
        will instead use the default implementation
        :class:`wtforms.csrf.session.SessionCSRF`.

        :param form: The form.
        :return: A CSRF implementation.
        """
        if self.csrf_class is not None:
            return self.csrf_class()

        from wtforms.csrf.session import SessionCSRF

        return SessionCSRF()

    # -- i18n

    locales = False
    cache_translations = True
    translations_cache = {}

    def get_translations(self, form):
        """
        Override in subclasses to provide alternate translations factory.
        See the i18n documentation for more.

        :param form: The form.
        :return: An object that provides gettext() and ngettext() methods.
        """
        locales = self.locales
        if locales is False:
            return None

        if self.cache_translations:
            # Make locales be a hashable value
            locales = tuple(locales) if locales else None

            translations = self.translations_cache.get(locales)
            if translations is None:
                translations = self.translations_cache[locales] = i18n.get_translations(
                    locales
                )

            return translations

        return i18n.get_translations(locales)

    # -- General

    def update_values(self, values):
        """
        Given a dictionary of values, update values on this `Meta` instance.
        """
        for key, value in values.items():
            setattr(self, key, value)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_core\tests\test_numeric.py ===
import itertools
import math
import platform
import sys
import warnings
from decimal import Decimal

import pytest
from hypothesis import given
from hypothesis import strategies as st
from hypothesis.extra import numpy as hynp
from numpy._core._rational_tests import rational

import numpy as np
from numpy import ma
from numpy._core import sctypes
from numpy._core.numerictypes import obj2sctype
from numpy.exceptions import AxisError
from numpy.random import rand, randint, randn
from numpy.testing import (
    HAS_REFCOUNT,
    IS_WASM,
    assert_,
    assert_almost_equal,
    assert_array_almost_equal,
    assert_array_equal,
    assert_array_max_ulp,
    assert_equal,
    assert_raises,
    assert_raises_regex,
)


class TestResize:
    def test_copies(self):
        A = np.array([[1, 2], [3, 4]])
        Ar1 = np.array([[1, 2, 3, 4], [1, 2, 3, 4]])
        assert_equal(np.resize(A, (2, 4)), Ar1)

        Ar2 = np.array([[1, 2], [3, 4], [1, 2], [3, 4]])
        assert_equal(np.resize(A, (4, 2)), Ar2)

        Ar3 = np.array([[1, 2, 3], [4, 1, 2], [3, 4, 1], [2, 3, 4]])
        assert_equal(np.resize(A, (4, 3)), Ar3)

    def test_repeats(self):
        A = np.array([1, 2, 3])
        Ar1 = np.array([[1, 2, 3, 1], [2, 3, 1, 2]])
        assert_equal(np.resize(A, (2, 4)), Ar1)

        Ar2 = np.array([[1, 2], [3, 1], [2, 3], [1, 2]])
        assert_equal(np.resize(A, (4, 2)), Ar2)

        Ar3 = np.array([[1, 2, 3], [1, 2, 3], [1, 2, 3], [1, 2, 3]])
        assert_equal(np.resize(A, (4, 3)), Ar3)

    def test_zeroresize(self):
        A = np.array([[1, 2], [3, 4]])
        Ar = np.resize(A, (0,))
        assert_array_equal(Ar, np.array([]))
        assert_equal(A.dtype, Ar.dtype)

        Ar = np.resize(A, (0, 2))
        assert_equal(Ar.shape, (0, 2))

        Ar = np.resize(A, (2, 0))
        assert_equal(Ar.shape, (2, 0))

    def test_reshape_from_zero(self):
        # See also gh-6740
        A = np.zeros(0, dtype=[('a', np.float32)])
        Ar = np.resize(A, (2, 1))
        assert_array_equal(Ar, np.zeros((2, 1), Ar.dtype))
        assert_equal(A.dtype, Ar.dtype)

    def test_negative_resize(self):
        A = np.arange(0, 10, dtype=np.float32)
        new_shape = (-10, -1)
        with pytest.raises(ValueError, match=r"negative"):
            np.resize(A, new_shape=new_shape)

    def test_subclass(self):
        class MyArray(np.ndarray):
            __array_priority__ = 1.

        my_arr = np.array([1]).view(MyArray)
        assert type(np.resize(my_arr, 5)) is MyArray
        assert type(np.resize(my_arr, 0)) is MyArray

        my_arr = np.array([]).view(MyArray)
        assert type(np.resize(my_arr, 5)) is MyArray


class TestNonarrayArgs:
    # check that non-array arguments to functions wrap them in arrays
    def test_choose(self):
        choices = [[0, 1, 2],
                   [3, 4, 5],
                   [5, 6, 7]]
        tgt = [5, 1, 5]
        a = [2, 0, 1]

        out = np.choose(a, choices)
        assert_equal(out, tgt)

    def test_clip(self):
        arr = [-1, 5, 2, 3, 10, -4, -9]
        out = np.clip(arr, 2, 7)
        tgt = [2, 5, 2, 3, 7, 2, 2]
        assert_equal(out, tgt)

    def test_compress(self):
        arr = [[0, 1, 2, 3, 4],
               [5, 6, 7, 8, 9]]
        tgt = [[5, 6, 7, 8, 9]]
        out = np.compress([0, 1], arr, axis=0)
        assert_equal(out, tgt)

    def test_count_nonzero(self):
        arr = [[0, 1, 7, 0, 0],
               [3, 0, 0, 2, 19]]
        tgt = np.array([2, 3])
        out = np.count_nonzero(arr, axis=1)
        assert_equal(out, tgt)

    def test_diagonal(self):
        a = [[0, 1, 2, 3],
             [4, 5, 6, 7],
             [8, 9, 10, 11]]
        out = np.diagonal(a)
        tgt = [0, 5, 10]

        assert_equal(out, tgt)

    def test_mean(self):
        A = [[1, 2, 3], [4, 5, 6]]
        assert_(np.mean(A) == 3.5)
        assert_(np.all(np.mean(A, 0) == np.array([2.5, 3.5, 4.5])))
        assert_(np.all(np.mean(A, 1) == np.array([2., 5.])))

        with warnings.catch_warnings(record=True) as w:
            warnings.filterwarnings('always', '', RuntimeWarning)
            assert_(np.isnan(np.mean([])))
            assert_(w[0].category is RuntimeWarning)

    def test_ptp(self):
        a = [3, 4, 5, 10, -3, -5, 6.0]
        assert_equal(np.ptp(a, axis=0), 15.0)

    def test_prod(self):
        arr = [[1, 2, 3, 4],
               [5, 6, 7, 9],
               [10, 3, 4, 5]]
        tgt = [24, 1890, 600]

        assert_equal(np.prod(arr, axis=-1), tgt)

    def test_ravel(self):
        a = [[1, 2, 3], [4, 5, 6], [7, 8, 9], [10, 11, 12]]
        tgt = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
        assert_equal(np.ravel(a), tgt)

    def test_repeat(self):
        a = [1, 2, 3]
        tgt = [1, 1, 2, 2, 3, 3]

        out = np.repeat(a, 2)
        assert_equal(out, tgt)

    def test_reshape(self):
        arr = [[1, 2, 3], [4, 5, 6], [7, 8, 9], [10, 11, 12]]
        tgt = [[1, 2, 3, 4, 5, 6], [7, 8, 9, 10, 11, 12]]
        assert_equal(np.reshape(arr, (2, 6)), tgt)

    def test_reshape_shape_arg(self):
        arr = np.arange(12)
        shape = (3, 4)
        expected = arr.reshape(shape)

        with pytest.raises(
            TypeError,
            match="You cannot specify 'newshape' and 'shape' "
                  "arguments at the same time."
        ):
            np.reshape(arr, shape=shape, newshape=shape)
        with pytest.raises(
            TypeError,
            match=r"reshape\(\) missing 1 required positional "
                  "argument: 'shape'"
        ):
            np.reshape(arr)

        assert_equal(np.reshape(arr, shape), expected)
        assert_equal(np.reshape(arr, shape, order="C"), expected)
        assert_equal(np.reshape(arr, shape, "C"), expected)
        assert_equal(np.reshape(arr, shape=shape), expected)
        assert_equal(np.reshape(arr, shape=shape, order="C"), expected)
        with pytest.warns(DeprecationWarning):
            actual = np.reshape(arr, newshape=shape)
            assert_equal(actual, expected)

    def test_reshape_copy_arg(self):
        arr = np.arange(24).reshape(2, 3, 4)
        arr_f_ord = np.array(arr, order="F")
        shape = (12, 2)

        assert np.shares_memory(np.reshape(arr, shape), arr)
        assert np.shares_memory(np.reshape(arr, shape, order="C"), arr)
        assert np.shares_memory(
            np.reshape(arr_f_ord, shape, order="F"), arr_f_ord)
        assert np.shares_memory(np.reshape(arr, shape, copy=None), arr)
        assert np.shares_memory(np.reshape(arr, shape, copy=False), arr)
        assert np.shares_memory(arr.reshape(shape, copy=False), arr)
        assert not np.shares_memory(np.reshape(arr, shape, copy=True), arr)
        assert not np.shares_memory(
            np.reshape(arr, shape, order="C", copy=True), arr)
        assert not np.shares_memory(
            np.reshape(arr, shape, order="F", copy=True), arr)
        assert not np.shares_memory(
            np.reshape(arr, shape, order="F", copy=None), arr)

        err_msg = "Unable to avoid creating a copy while reshaping."
        with pytest.raises(ValueError, match=err_msg):
            np.reshape(arr, shape, order="F", copy=False)
        with pytest.raises(ValueError, match=err_msg):
            np.reshape(arr_f_ord, shape, order="C", copy=False)

    def test_round(self):
        arr = [1.56, 72.54, 6.35, 3.25]
        tgt = [1.6, 72.5, 6.4, 3.2]
        assert_equal(np.around(arr, decimals=1), tgt)
        s = np.float64(1.)
        assert_(isinstance(s.round(), np.float64))
        assert_equal(s.round(), 1.)

    @pytest.mark.parametrize('dtype', [
        np.int8, np.int16, np.int32, np.int64,
        np.uint8, np.uint16, np.uint32, np.uint64,
        np.float16, np.float32, np.float64,
    ])
    def test_dunder_round(self, dtype):
        s = dtype(1)
        assert_(isinstance(round(s), int))
        assert_(isinstance(round(s, None), int))
        assert_(isinstance(round(s, ndigits=None), int))
        assert_equal(round(s), 1)
        assert_equal(round(s, None), 1)
        assert_equal(round(s, ndigits=None), 1)

    @pytest.mark.parametrize('val, ndigits', [
        pytest.param(2**31 - 1, -1,
            marks=pytest.mark.skip(reason="Out of range of int32")
        ),
        (2**31 - 1, 1 - math.ceil(math.log10(2**31 - 1))),
        (2**31 - 1, -math.ceil(math.log10(2**31 - 1)))
    ])
    def test_dunder_round_edgecases(self, val, ndigits):
        assert_equal(round(val, ndigits), round(np.int32(val), ndigits))

    def test_dunder_round_accuracy(self):
        f = np.float64(5.1 * 10**73)
        assert_(isinstance(round(f, -73), np.float64))
        assert_array_max_ulp(round(f, -73), 5.0 * 10**73)
        assert_(isinstance(round(f, ndigits=-73), np.float64))
        assert_array_max_ulp(round(f, ndigits=-73), 5.0 * 10**73)

        i = np.int64(501)
        assert_(isinstance(round(i, -2), np.int64))
        assert_array_max_ulp(round(i, -2), 500)
        assert_(isinstance(round(i, ndigits=-2), np.int64))
        assert_array_max_ulp(round(i, ndigits=-2), 500)

    @pytest.mark.xfail(raises=AssertionError, reason="gh-15896")
    def test_round_py_consistency(self):
        f = 5.1 * 10**73
        assert_equal(round(np.float64(f), -73), round(f, -73))

    def test_searchsorted(self):
        arr = [-8, -5, -1, 3, 6, 10]
        out = np.searchsorted(arr, 0)
        assert_equal(out, 3)

    def test_size(self):
        A = [[1, 2, 3], [4, 5, 6]]
        assert_(np.size(A) == 6)
        assert_(np.size(A, 0) == 2)
        assert_(np.size(A, 1) == 3)

    def test_squeeze(self):
        A = [[[1, 1, 1], [2, 2, 2], [3, 3, 3]]]
        assert_equal(np.squeeze(A).shape, (3, 3))
        assert_equal(np.squeeze(np.zeros((1, 3, 1))).shape, (3,))
        assert_equal(np.squeeze(np.zeros((1, 3, 1)), axis=0).shape, (3, 1))
        assert_equal(np.squeeze(np.zeros((1, 3, 1)), axis=-1).shape, (1, 3))
        assert_equal(np.squeeze(np.zeros((1, 3, 1)), axis=2).shape, (1, 3))
        assert_equal(np.squeeze([np.zeros((3, 1))]).shape, (3,))
        assert_equal(np.squeeze([np.zeros((3, 1))], axis=0).shape, (3, 1))
        assert_equal(np.squeeze([np.zeros((3, 1))], axis=2).shape, (1, 3))
        assert_equal(np.squeeze([np.zeros((3, 1))], axis=-1).shape, (1, 3))

    def test_std(self):
        A = [[1, 2, 3], [4, 5, 6]]
        assert_almost_equal(np.std(A), 1.707825127659933)
        assert_almost_equal(np.std(A, 0), np.array([1.5, 1.5, 1.5]))
        assert_almost_equal(np.std(A, 1), np.array([0.81649658, 0.81649658]))

        with warnings.catch_warnings(record=True) as w:
            warnings.filterwarnings('always', '', RuntimeWarning)
            assert_(np.isnan(np.std([])))
            assert_(w[0].category is RuntimeWarning)

    def test_swapaxes(self):
        tgt = [[[0, 4], [2, 6]], [[1, 5], [3, 7]]]
        a = [[[0, 1], [2, 3]], [[4, 5], [6, 7]]]
        out = np.swapaxes(a, 0, 2)
        assert_equal(out, tgt)

    def test_sum(self):
        m = [[1, 2, 3],
             [4, 5, 6],
             [7, 8, 9]]
        tgt = [[6], [15], [24]]
        out = np.sum(m, axis=1, keepdims=True)

        assert_equal(tgt, out)

    def test_take(self):
        tgt = [2, 3, 5]
        indices = [1, 2, 4]
        a = [1, 2, 3, 4, 5]

        out = np.take(a, indices)
        assert_equal(out, tgt)

        pairs = [
            (np.int32, np.int32), (np.int32, np.int64),
            (np.int64, np.int32), (np.int64, np.int64)
        ]
        for array_type, indices_type in pairs:
            x = np.array([1, 2, 3, 4, 5], dtype=array_type)
            ind = np.array([0, 2, 2, 3], dtype=indices_type)
            tgt = np.array([1, 3, 3, 4], dtype=array_type)
            out = np.take(x, ind)
            assert_equal(out, tgt)
            assert_equal(out.dtype, tgt.dtype)

    def test_trace(self):
        c = [[1, 2], [3, 4], [5, 6]]
        assert_equal(np.trace(c), 5)

    def test_transpose(self):
        arr = [[1, 2], [3, 4], [5, 6]]
        tgt = [[1, 3, 5], [2, 4, 6]]
        assert_equal(np.transpose(arr, (1, 0)), tgt)
        assert_equal(np.transpose(arr, (-1, -2)), tgt)
        assert_equal(np.matrix_transpose(arr), tgt)

    def test_var(self):
        A = [[1, 2, 3], [4, 5, 6]]
        assert_almost_equal(np.var(A), 2.9166666666666665)
        assert_almost_equal(np.var(A, 0), np.array([2.25, 2.25, 2.25]))
        assert_almost_equal(np.var(A, 1), np.array([0.66666667, 0.66666667]))

        with warnings.catch_warnings(record=True) as w:
            warnings.filterwarnings('always', '', RuntimeWarning)
            assert_(np.isnan(np.var([])))
            assert_(w[0].category is RuntimeWarning)

        B = np.array([None, 0])
        B[0] = 1j
        assert_almost_equal(np.var(B), 0.25)

    def test_std_with_mean_keyword(self):
        # Setting the seed to make the test reproducible
        rng = np.random.RandomState(1234)
        A = rng.randn(10, 20, 5) + 0.5

        mean_out = np.zeros((10, 1, 5))
        std_out = np.zeros((10, 1, 5))

        mean = np.mean(A,
                       out=mean_out,
                       axis=1,
                       keepdims=True)

        # The returned  object should be the object specified during calling
        assert mean_out is mean

        std = np.std(A,
                     out=std_out,
                     axis=1,
                     keepdims=True,
                     mean=mean)

        # The returned  object should be the object specified during calling
        assert std_out is std

        # Shape of returned mean and std should be same
        assert std.shape == mean.shape
        assert std.shape == (10, 1, 5)

        # Output should be the same as from the individual algorithms
        std_old = np.std(A, axis=1, keepdims=True)

        assert std_old.shape == mean.shape
        assert_almost_equal(std, std_old)

    def test_var_with_mean_keyword(self):
        # Setting the seed to make the test reproducible
        rng = np.random.RandomState(1234)
        A = rng.randn(10, 20, 5) + 0.5

        mean_out = np.zeros((10, 1, 5))
        var_out = np.zeros((10, 1, 5))

        mean = np.mean(A,
                       out=mean_out,
                       axis=1,
                       keepdims=True)

        # The returned  object should be the object specified during calling
        assert mean_out is mean

        var = np.var(A,
                     out=var_out,
                     axis=1,
                     keepdims=True,
                     mean=mean)

        # The returned  object should be the object specified during calling
        assert var_out is var

        # Shape of returned mean and var should be same
        assert var.shape == mean.shape
        assert var.shape == (10, 1, 5)

        # Output should be the same as from the individual algorithms
        var_old = np.var(A, axis=1, keepdims=True)

        assert var_old.shape == mean.shape
        assert_almost_equal(var, var_old)

    def test_std_with_mean_keyword_keepdims_false(self):
        rng = np.random.RandomState(1234)
        A = rng.randn(10, 20, 5) + 0.5

        mean = np.mean(A,
                       axis=1,
                       keepdims=True)

        std = np.std(A,
                     axis=1,
                     keepdims=False,
                     mean=mean)

        # Shape of returned mean and std should be same
        assert std.shape == (10, 5)

        # Output should be the same as from the individual algorithms
        std_old = np.std(A, axis=1, keepdims=False)
        mean_old = np.mean(A, axis=1, keepdims=False)

        assert std_old.shape == mean_old.shape
        assert_equal(std, std_old)

    def test_var_with_mean_keyword_keepdims_false(self):
        rng = np.random.RandomState(1234)
        A = rng.randn(10, 20, 5) + 0.5

        mean = np.mean(A,
                       axis=1,
                       keepdims=True)

        var = np.var(A,
                     axis=1,
                     keepdims=False,
                     mean=mean)

        # Shape of returned mean and var should be same
        assert var.shape == (10, 5)

        # Output should be the same as from the individual algorithms
        var_old = np.var(A, axis=1, keepdims=False)
        mean_old = np.mean(A, axis=1, keepdims=False)

        assert var_old.shape == mean_old.shape
        assert_equal(var, var_old)

    def test_std_with_mean_keyword_where_nontrivial(self):
        rng = np.random.RandomState(1234)
        A = rng.randn(10, 20, 5) + 0.5

        where = A > 0.5

        mean = np.mean(A,
                       axis=1,
                       keepdims=True,
                       where=where)

        std = np.std(A,
                     axis=1,
                     keepdims=False,
                     mean=mean,
                     where=where)

        # Shape of returned mean and std should be same
        assert std.shape == (10, 5)

        # Output should be the same as from the individual algorithms
        std_old = np.std(A, axis=1, where=where)
        mean_old = np.mean(A, axis=1, where=where)

        assert std_old.shape == mean_old.shape
        assert_equal(std, std_old)

    def test_var_with_mean_keyword_where_nontrivial(self):
        rng = np.random.RandomState(1234)
        A = rng.randn(10, 20, 5) + 0.5

        where = A > 0.5

        mean = np.mean(A,
                       axis=1,
                       keepdims=True,
                       where=where)

        var = np.var(A,
                     axis=1,
                     keepdims=False,
                     mean=mean,
                     where=where)

        # Shape of returned mean and var should be same
        assert var.shape == (10, 5)

        # Output should be the same as from the individual algorithms
        var_old = np.var(A, axis=1, where=where)
        mean_old = np.mean(A, axis=1, where=where)

        assert var_old.shape == mean_old.shape
        assert_equal(var, var_old)

    def test_std_with_mean_keyword_multiple_axis(self):
        # Setting the seed to make the test reproducible
        rng = np.random.RandomState(1234)
        A = rng.randn(10, 20, 5) + 0.5

        axis = (0, 2)

        mean = np.mean(A,
                       out=None,
                       axis=axis,
                       keepdims=True)

        std = np.std(A,
                     out=None,
                     axis=axis,
                     keepdims=False,
                     mean=mean)

        # Shape of returned mean and std should be same
        assert std.shape == (20,)

        # Output should be the same as from the individual algorithms
        std_old = np.std(A, axis=axis, keepdims=False)

        assert_almost_equal(std, std_old)

    def test_std_with_mean_keyword_axis_None(self):
        # Setting the seed to make the test reproducible
        rng = np.random.RandomState(1234)
        A = rng.randn(10, 20, 5) + 0.5

        axis = None

        mean = np.mean(A,
                       out=None,
                       axis=axis,
                       keepdims=True)

        std = np.std(A,
                     out=None,
                     axis=axis,
                     keepdims=False,
                     mean=mean)

        # Shape of returned mean and std should be same
        assert std.shape == ()

        # Output should be the same as from the individual algorithms
        std_old = np.std(A, axis=axis, keepdims=False)

        assert_almost_equal(std, std_old)

    def test_std_with_mean_keyword_keepdims_true_masked(self):

        A = ma.array([[2., 3., 4., 5.],
                      [1., 2., 3., 4.]],
                     mask=[[True, False, True, False],
                           [True, False, True, False]])

        B = ma.array([[100., 3., 104., 5.],
                      [101., 2., 103., 4.]],
                      mask=[[True, False, True, False],
                            [True, False, True, False]])

        mean_out = ma.array([[0., 0., 0., 0.]],
                            mask=[[False, False, False, False]])
        std_out = ma.array([[0., 0., 0., 0.]],
                           mask=[[False, False, False, False]])

        axis = 0

        mean = np.mean(A, out=mean_out,
                       axis=axis, keepdims=True)

        std = np.std(A, out=std_out,
                     axis=axis, keepdims=True,
                     mean=mean)

        # Shape of returned mean and std should be same
        assert std.shape == mean.shape
        assert std.shape == (1, 4)

        # Output should be the same as from the individual algorithms
        std_old = np.std(A, axis=axis, keepdims=True)
        mean_old = np.mean(A, axis=axis, keepdims=True)

        assert std_old.shape == mean_old.shape
        assert_almost_equal(std, std_old)
        assert_almost_equal(mean, mean_old)

        assert mean_out is mean
        assert std_out is std

        # masked elements should be ignored
        mean_b = np.mean(B, axis=axis, keepdims=True)
        std_b = np.std(B, axis=axis, keepdims=True, mean=mean_b)
        assert_almost_equal(std, std_b)
        assert_almost_equal(mean, mean_b)

    def test_var_with_mean_keyword_keepdims_true_masked(self):

        A = ma.array([[2., 3., 4., 5.],
                      [1., 2., 3., 4.]],
                     mask=[[True, False, True, False],
                           [True, False, True, False]])

        B = ma.array([[100., 3., 104., 5.],
                      [101., 2., 103., 4.]],
                      mask=[[True, False, True, False],
                            [True, False, True, False]])

        mean_out = ma.array([[0., 0., 0., 0.]],
                            mask=[[False, False, False, False]])
        var_out = ma.array([[0., 0., 0., 0.]],
                           mask=[[False, False, False, False]])

        axis = 0

        mean = np.mean(A, out=mean_out,
                       axis=axis, keepdims=True)

        var = np.var(A, out=var_out,
                     axis=axis, keepdims=True,
                     mean=mean)

        # Shape of returned mean and var should be same
        assert var.shape == mean.shape
        assert var.shape == (1, 4)

        # Output should be the same as from the individual algorithms
        var_old = np.var(A, axis=axis, keepdims=True)
        mean_old = np.mean(A, axis=axis, keepdims=True)

        assert var_old.shape == mean_old.shape
        assert_almost_equal(var, var_old)
        assert_almost_equal(mean, mean_old)

        assert mean_out is mean
        assert var_out is var

        # masked elements should be ignored
        mean_b = np.mean(B, axis=axis, keepdims=True)
        var_b = np.var(B, axis=axis, keepdims=True, mean=mean_b)
        assert_almost_equal(var, var_b)
        assert_almost_equal(mean, mean_b)


class TestIsscalar:
    def test_isscalar(self):
        assert_(np.isscalar(3.1))
        assert_(np.isscalar(np.int16(12345)))
        assert_(np.isscalar(False))
        assert_(np.isscalar('numpy'))
        assert_(not np.isscalar([3.1]))
        assert_(not np.isscalar(None))

        # PEP 3141
        from fractions import Fraction
        assert_(np.isscalar(Fraction(5, 17)))
        from numbers import Number
        assert_(np.isscalar(Number()))


class TestBoolScalar:
    def test_logical(self):
        f = np.False_
        t = np.True_
        s = "xyz"
        assert_((t and s) is s)
        assert_((f and s) is f)

    def test_bitwise_or(self):
        f = np.False_
        t = np.True_
        assert_((t | t) is t)
        assert_((f | t) is t)
        assert_((t | f) is t)
        assert_((f | f) is f)

    def test_bitwise_and(self):
        f = np.False_
        t = np.True_
        assert_((t & t) is t)
        assert_((f & t) is f)
        assert_((t & f) is f)
        assert_((f & f) is f)

    def test_bitwise_xor(self):
        f = np.False_
        t = np.True_
        assert_((t ^ t) is f)
        assert_((f ^ t) is t)
        assert_((t ^ f) is t)
        assert_((f ^ f) is f)


class TestBoolArray:
    def setup_method(self):
        # offset for simd tests
        self.t = np.array([True] * 41, dtype=bool)[1::]
        self.f = np.array([False] * 41, dtype=bool)[1::]
        self.o = np.array([False] * 42, dtype=bool)[2::]
        self.nm = self.f.copy()
        self.im = self.t.copy()
        self.nm[3] = True
        self.nm[-2] = True
        self.im[3] = False
        self.im[-2] = False

    def test_all_any(self):
        assert_(self.t.all())
        assert_(self.t.any())
        assert_(not self.f.all())
        assert_(not self.f.any())
        assert_(self.nm.any())
        assert_(self.im.any())
        assert_(not self.nm.all())
        assert_(not self.im.all())
        # check bad element in all positions
        for i in range(256 - 7):
            d = np.array([False] * 256, dtype=bool)[7::]
            d[i] = True
            assert_(np.any(d))
            e = np.array([True] * 256, dtype=bool)[7::]
            e[i] = False
            assert_(not np.all(e))
            assert_array_equal(e, ~d)
        # big array test for blocked libc loops
        for i in list(range(9, 6000, 507)) + [7764, 90021, -10]:
            d = np.array([False] * 100043, dtype=bool)
            d[i] = True
            assert_(np.any(d), msg=f"{i!r}")
            e = np.array([True] * 100043, dtype=bool)
            e[i] = False
            assert_(not np.all(e), msg=f"{i!r}")

    def test_logical_not_abs(self):
        assert_array_equal(~self.t, self.f)
        assert_array_equal(np.abs(~self.t), self.f)
        assert_array_equal(np.abs(~self.f), self.t)
        assert_array_equal(np.abs(self.f), self.f)
        assert_array_equal(~np.abs(self.f), self.t)
        assert_array_equal(~np.abs(self.t), self.f)
        assert_array_equal(np.abs(~self.nm), self.im)
        np.logical_not(self.t, out=self.o)
        assert_array_equal(self.o, self.f)
        np.abs(self.t, out=self.o)
        assert_array_equal(self.o, self.t)

    def test_logical_and_or_xor(self):
        assert_array_equal(self.t | self.t, self.t)
        assert_array_equal(self.f | self.f, self.f)
        assert_array_equal(self.t | self.f, self.t)
        assert_array_equal(self.f | self.t, self.t)
        np.logical_or(self.t, self.t, out=self.o)
        assert_array_equal(self.o, self.t)
        assert_array_equal(self.t & self.t, self.t)
        assert_array_equal(self.f & self.f, self.f)
        assert_array_equal(self.t & self.f, self.f)
        assert_array_equal(self.f & self.t, self.f)
        np.logical_and(self.t, self.t, out=self.o)
        assert_array_equal(self.o, self.t)
        assert_array_equal(self.t ^ self.t, self.f)
        assert_array_equal(self.f ^ self.f, self.f)
        assert_array_equal(self.t ^ self.f, self.t)
        assert_array_equal(self.f ^ self.t, self.t)
        np.logical_xor(self.t, self.t, out=self.o)
        assert_array_equal(self.o, self.f)

        assert_array_equal(self.nm & self.t, self.nm)
        assert_array_equal(self.im & self.f, False)
        assert_array_equal(self.nm & True, self.nm)
        assert_array_equal(self.im & False, self.f)
        assert_array_equal(self.nm | self.t, self.t)
        assert_array_equal(self.im | self.f, self.im)
        assert_array_equal(self.nm | True, self.t)
        assert_array_equal(self.im | False, self.im)
        assert_array_equal(self.nm ^ self.t, self.im)
        assert_array_equal(self.im ^ self.f, self.im)
        assert_array_equal(self.nm ^ True, self.im)
        assert_array_equal(self.im ^ False, self.im)


class TestBoolCmp:
    def setup_method(self):
        self.f = np.ones(256, dtype=np.float32)
        self.ef = np.ones(self.f.size, dtype=bool)
        self.d = np.ones(128, dtype=np.float64)
        self.ed = np.ones(self.d.size, dtype=bool)
        # generate values for all permutation of 256bit simd vectors
        s = 0
        for i in range(32):
            self.f[s:s + 8] = [i & 2**x for x in range(8)]
            self.ef[s:s + 8] = [(i & 2**x) != 0 for x in range(8)]
            s += 8
        s = 0
        for i in range(16):
            self.d[s:s + 4] = [i & 2**x for x in range(4)]
            self.ed[s:s + 4] = [(i & 2**x) != 0 for x in range(4)]
            s += 4

        self.nf = self.f.copy()
        self.nd = self.d.copy()
        self.nf[self.ef] = np.nan
        self.nd[self.ed] = np.nan

        self.inff = self.f.copy()
        self.infd = self.d.copy()
        self.inff[::3][self.ef[::3]] = np.inf
        self.infd[::3][self.ed[::3]] = np.inf
        self.inff[1::3][self.ef[1::3]] = -np.inf
        self.infd[1::3][self.ed[1::3]] = -np.inf
        self.inff[2::3][self.ef[2::3]] = np.nan
        self.infd[2::3][self.ed[2::3]] = np.nan
        self.efnonan = self.ef.copy()
        self.efnonan[2::3] = False
        self.ednonan = self.ed.copy()
        self.ednonan[2::3] = False

        self.signf = self.f.copy()
        self.signd = self.d.copy()
        self.signf[self.ef] *= -1.
        self.signd[self.ed] *= -1.
        self.signf[1::6][self.ef[1::6]] = -np.inf
        self.signd[1::6][self.ed[1::6]] = -np.inf
        # On RISC-V, many operations that produce NaNs, such as converting
        # a -NaN from f64 to f32, return a canonical NaN.  The canonical
        # NaNs are always positive.  See section 11.3 NaN Generation and
        # Propagation of the RISC-V Unprivileged ISA for more details.
        # We disable the float32 sign test on riscv64 for -np.nan as the sign
        # of the NaN will be lost when it's converted to a float32.
        if platform.machine() != 'riscv64':
            self.signf[3::6][self.ef[3::6]] = -np.nan
        self.signd[3::6][self.ed[3::6]] = -np.nan
        self.signf[4::6][self.ef[4::6]] = -0.
        self.signd[4::6][self.ed[4::6]] = -0.

    def test_float(self):
        # offset for alignment test
        for i in range(4):
            assert_array_equal(self.f[i:] > 0, self.ef[i:])
            assert_array_equal(self.f[i:] - 1 >= 0, self.ef[i:])
            assert_array_equal(self.f[i:] == 0, ~self.ef[i:])
            assert_array_equal(-self.f[i:] < 0, self.ef[i:])
            assert_array_equal(-self.f[i:] + 1 <= 0, self.ef[i:])
            r = self.f[i:] != 0
            assert_array_equal(r, self.ef[i:])
            r2 = self.f[i:] != np.zeros_like(self.f[i:])
            r3 = 0 != self.f[i:]
            assert_array_equal(r, r2)
            assert_array_equal(r, r3)
            # check bool == 0x1
            assert_array_equal(r.view(np.int8), r.astype(np.int8))
            assert_array_equal(r2.view(np.int8), r2.astype(np.int8))
            assert_array_equal(r3.view(np.int8), r3.astype(np.int8))

            # isnan on amd64 takes the same code path
            assert_array_equal(np.isnan(self.nf[i:]), self.ef[i:])
            assert_array_equal(np.isfinite(self.nf[i:]), ~self.ef[i:])
            assert_array_equal(np.isfinite(self.inff[i:]), ~self.ef[i:])
            assert_array_equal(np.isinf(self.inff[i:]), self.efnonan[i:])
            assert_array_equal(np.signbit(self.signf[i:]), self.ef[i:])

    def test_double(self):
        # offset for alignment test
        for i in range(2):
            assert_array_equal(self.d[i:] > 0, self.ed[i:])
            assert_array_equal(self.d[i:] - 1 >= 0, self.ed[i:])
            assert_array_equal(self.d[i:] == 0, ~self.ed[i:])
            assert_array_equal(-self.d[i:] < 0, self.ed[i:])
            assert_array_equal(-self.d[i:] + 1 <= 0, self.ed[i:])
            r = self.d[i:] != 0
            assert_array_equal(r, self.ed[i:])
            r2 = self.d[i:] != np.zeros_like(self.d[i:])
            r3 = 0 != self.d[i:]
            assert_array_equal(r, r2)
            assert_array_equal(r, r3)
            # check bool == 0x1
            assert_array_equal(r.view(np.int8), r.astype(np.int8))
            assert_array_equal(r2.view(np.int8), r2.astype(np.int8))
            assert_array_equal(r3.view(np.int8), r3.astype(np.int8))

            # isnan on amd64 takes the same code path
            assert_array_equal(np.isnan(self.nd[i:]), self.ed[i:])
            assert_array_equal(np.isfinite(self.nd[i:]), ~self.ed[i:])
            assert_array_equal(np.isfinite(self.infd[i:]), ~self.ed[i:])
            assert_array_equal(np.isinf(self.infd[i:]), self.ednonan[i:])
            assert_array_equal(np.signbit(self.signd[i:]), self.ed[i:])


class TestSeterr:
    def test_default(self):
        err = np.geterr()
        assert_equal(err,
                     {'divide': 'warn',
                          'invalid': 'warn',
                          'over': 'warn',
                          'under': 'ignore'}
                     )

    def test_set(self):
        with np.errstate():
            err = np.seterr()
            old = np.seterr(divide='print')
            assert_(err == old)
            new = np.seterr()
            assert_(new['divide'] == 'print')
            np.seterr(over='raise')
            assert_(np.geterr()['over'] == 'raise')
            assert_(new['divide'] == 'print')
            np.seterr(**old)
            assert_(np.geterr() == old)

    @pytest.mark.skipif(IS_WASM, reason="no wasm fp exception support")
    @pytest.mark.skipif(platform.machine() == "armv5tel", reason="See gh-413.")
    def test_divide_err(self):
        with np.errstate(divide='raise'):
            with assert_raises(FloatingPointError):
                np.array([1.]) / np.array([0.])

            np.seterr(divide='ignore')
            np.array([1.]) / np.array([0.])


class TestFloatExceptions:
    def assert_raises_fpe(self, fpeerr, flop, x, y):
        ftype = type(x)
        try:
            flop(x, y)
            assert_(False,
                    f"Type {ftype} did not raise fpe error '{fpeerr}'.")
        except FloatingPointError as exc:
            assert_(str(exc).find(fpeerr) >= 0,
                    f"Type {ftype} raised wrong fpe error '{exc}'.")

    def assert_op_raises_fpe(self, fpeerr, flop, sc1, sc2):
        # Check that fpe exception is raised.
        #
        # Given a floating operation `flop` and two scalar values, check that
        # the operation raises the floating point exception specified by
        # `fpeerr`. Tests all variants with 0-d array scalars as well.

        self.assert_raises_fpe(fpeerr, flop, sc1, sc2)
        self.assert_raises_fpe(fpeerr, flop, sc1[()], sc2)
        self.assert_raises_fpe(fpeerr, flop, sc1, sc2[()])
        self.assert_raises_fpe(fpeerr, flop, sc1[()], sc2[()])

    # Test for all real and complex float types
    @pytest.mark.skipif(IS_WASM, reason="no wasm fp exception support")
    @pytest.mark.parametrize("typecode", np.typecodes["AllFloat"])
    def test_floating_exceptions(self, typecode):
        if 'bsd' in sys.platform and typecode in 'gG':
            pytest.skip(reason="Fallback impl for (c)longdouble may not raise "
                               "FPE errors as expected on BSD OSes, "
                               "see gh-24876, gh-23379")

        # Test basic arithmetic function errors
        with np.errstate(all='raise'):
            ftype = obj2sctype(typecode)
            if np.dtype(ftype).kind == 'f':
                # Get some extreme values for the type
                fi = np.finfo(ftype)
                ft_tiny = fi._machar.tiny
                ft_max = fi.max
                ft_eps = fi.eps
                underflow = 'underflow'
                divbyzero = 'divide by zero'
            else:
                # 'c', complex, corresponding real dtype
                rtype = type(ftype(0).real)
                fi = np.finfo(rtype)
                ft_tiny = ftype(fi._machar.tiny)
                ft_max = ftype(fi.max)
                ft_eps = ftype(fi.eps)
                # The complex types raise different exceptions
                underflow = ''
                divbyzero = ''
            overflow = 'overflow'
            invalid = 'invalid'

            # The value of tiny for double double is NaN, so we need to
            # pass the assert
            if not np.isnan(ft_tiny):
                self.assert_raises_fpe(underflow,
                                    lambda a, b: a / b, ft_tiny, ft_max)
                self.assert_raises_fpe(underflow,
                                    lambda a, b: a * b, ft_tiny, ft_tiny)
            self.assert_raises_fpe(overflow,
                                   lambda a, b: a * b, ft_max, ftype(2))
            self.assert_raises_fpe(overflow,
                                   lambda a, b: a / b, ft_max, ftype(0.5))
            self.assert_raises_fpe(overflow,
                                   lambda a, b: a + b, ft_max, ft_max * ft_eps)
            self.assert_raises_fpe(overflow,
                                   lambda a, b: a - b, -ft_max, ft_max * ft_eps)
            self.assert_raises_fpe(overflow,
                                   np.power, ftype(2), ftype(2**fi.nexp))
            self.assert_raises_fpe(divbyzero,
                                   lambda a, b: a / b, ftype(1), ftype(0))
            self.assert_raises_fpe(
                invalid, lambda a, b: a / b, ftype(np.inf), ftype(np.inf)
            )
            self.assert_raises_fpe(invalid,
                                   lambda a, b: a / b, ftype(0), ftype(0))
            self.assert_raises_fpe(
                invalid, lambda a, b: a - b, ftype(np.inf), ftype(np.inf)
            )
            self.assert_raises_fpe(
                invalid, lambda a, b: a + b, ftype(np.inf), ftype(-np.inf)
            )
            self.assert_raises_fpe(invalid,
                                   lambda a, b: a * b, ftype(0), ftype(np.inf))

    @pytest.mark.skipif(IS_WASM, reason="no wasm fp exception support")
    def test_warnings(self):
        # test warning code path
        with warnings.catch_warnings(record=True) as w:
            warnings.simplefilter("always")
            with np.errstate(all="warn"):
                np.divide(1, 0.)
                assert_equal(len(w), 1)
                assert_("divide by zero" in str(w[0].message))
                np.array(1e300) * np.array(1e300)
                assert_equal(len(w), 2)
                assert_("overflow" in str(w[-1].message))
                np.array(np.inf) - np.array(np.inf)
                assert_equal(len(w), 3)
                assert_("invalid value" in str(w[-1].message))
                np.array(1e-300) * np.array(1e-300)
                assert_equal(len(w), 4)
                assert_("underflow" in str(w[-1].message))


class TestTypes:
    def check_promotion_cases(self, promote_func):
        # tests that the scalars get coerced correctly.
        b = np.bool(0)
        i8, i16, i32, i64 = np.int8(0), np.int16(0), np.int32(0), np.int64(0)
        u8, u16, u32, u64 = np.uint8(0), np.uint16(0), np.uint32(0), np.uint64(0)
        f32, f64, fld = np.float32(0), np.float64(0), np.longdouble(0)
        c64, c128, cld = np.complex64(0), np.complex128(0), np.clongdouble(0)

        # coercion within the same kind
        assert_equal(promote_func(i8, i16), np.dtype(np.int16))
        assert_equal(promote_func(i32, i8), np.dtype(np.int32))
        assert_equal(promote_func(i16, i64), np.dtype(np.int64))
        assert_equal(promote_func(u8, u32), np.dtype(np.uint32))
        assert_equal(promote_func(f32, f64), np.dtype(np.float64))
        assert_equal(promote_func(fld, f32), np.dtype(np.longdouble))
        assert_equal(promote_func(f64, fld), np.dtype(np.longdouble))
        assert_equal(promote_func(c128, c64), np.dtype(np.complex128))
        assert_equal(promote_func(cld, c128), np.dtype(np.clongdouble))
        assert_equal(promote_func(c64, fld), np.dtype(np.clongdouble))

        # coercion between kinds
        assert_equal(promote_func(b, i32), np.dtype(np.int32))
        assert_equal(promote_func(b, u8), np.dtype(np.uint8))
        assert_equal(promote_func(i8, u8), np.dtype(np.int16))
        assert_equal(promote_func(u8, i32), np.dtype(np.int32))
        assert_equal(promote_func(i64, u32), np.dtype(np.int64))
        assert_equal(promote_func(u64, i32), np.dtype(np.float64))
        assert_equal(promote_func(i32, f32), np.dtype(np.float64))
        assert_equal(promote_func(i64, f32), np.dtype(np.float64))
        assert_equal(promote_func(f32, i16), np.dtype(np.float32))
        assert_equal(promote_func(f32, u32), np.dtype(np.float64))
        assert_equal(promote_func(f32, c64), np.dtype(np.complex64))
        assert_equal(promote_func(c128, f32), np.dtype(np.complex128))
        assert_equal(promote_func(cld, f64), np.dtype(np.clongdouble))

        # coercion between scalars and 1-D arrays
        assert_equal(promote_func(np.array([b]), i8), np.dtype(np.int8))
        assert_equal(promote_func(np.array([b]), u8), np.dtype(np.uint8))
        assert_equal(promote_func(np.array([b]), i32), np.dtype(np.int32))
        assert_equal(promote_func(np.array([b]), u32), np.dtype(np.uint32))
        assert_equal(promote_func(np.array([i8]), i64), np.dtype(np.int64))
        # unsigned and signed unfortunately tend to promote to float64:
        assert_equal(promote_func(u64, np.array([i32])), np.dtype(np.float64))
        assert_equal(promote_func(i64, np.array([u32])), np.dtype(np.int64))
        assert_equal(promote_func(np.array([u16]), i32), np.dtype(np.int32))
        assert_equal(promote_func(np.int32(-1), np.array([u64])),
                     np.dtype(np.float64))
        assert_equal(promote_func(f64, np.array([f32])), np.dtype(np.float64))
        assert_equal(promote_func(fld, np.array([f32])),
                     np.dtype(np.longdouble))
        assert_equal(promote_func(np.array([f64]), fld),
                     np.dtype(np.longdouble))
        assert_equal(promote_func(fld, np.array([c64])),
                     np.dtype(np.clongdouble))
        assert_equal(promote_func(c64, np.array([f64])),
                     np.dtype(np.complex128))
        assert_equal(promote_func(np.complex64(3j), np.array([f64])),
                     np.dtype(np.complex128))
        assert_equal(promote_func(np.array([f32]), c128),
                     np.dtype(np.complex128))

        # coercion between scalars and 1-D arrays, where
        # the scalar has greater kind than the array
        assert_equal(promote_func(np.array([b]), f64), np.dtype(np.float64))
        assert_equal(promote_func(np.array([b]), i64), np.dtype(np.int64))
        assert_equal(promote_func(np.array([b]), u64), np.dtype(np.uint64))
        assert_equal(promote_func(np.array([i8]), f64), np.dtype(np.float64))
        assert_equal(promote_func(np.array([u16]), f64), np.dtype(np.float64))

    def test_coercion(self):
        def res_type(a, b):
            return np.add(a, b).dtype

        self.check_promotion_cases(res_type)

        # Use-case: float/complex scalar * bool/int8 array
        #           shouldn't narrow the float/complex type
        for a in [np.array([True, False]), np.array([-3, 12], dtype=np.int8)]:
            b = 1.234 * a
            assert_equal(b.dtype, np.dtype('f8'), f"array type {a.dtype}")
            b = np.longdouble(1.234) * a
            assert_equal(b.dtype, np.dtype(np.longdouble),
                         f"array type {a.dtype}")
            b = np.float64(1.234) * a
            assert_equal(b.dtype, np.dtype('f8'), f"array type {a.dtype}")
            b = np.float32(1.234) * a
            assert_equal(b.dtype, np.dtype('f4'), f"array type {a.dtype}")
            b = np.float16(1.234) * a
            assert_equal(b.dtype, np.dtype('f2'), f"array type {a.dtype}")

            b = 1.234j * a
            assert_equal(b.dtype, np.dtype('c16'), f"array type {a.dtype}")
            b = np.clongdouble(1.234j) * a
            assert_equal(b.dtype, np.dtype(np.clongdouble),
                         f"array type {a.dtype}")
            b = np.complex128(1.234j) * a
            assert_equal(b.dtype, np.dtype('c16'), f"array type {a.dtype}")
            b = np.complex64(1.234j) * a
            assert_equal(b.dtype, np.dtype('c8'), f"array type {a.dtype}")

        # The following use-case is problematic, and to resolve its
        # tricky side-effects requires more changes.
        #
        # Use-case: (1-t)*a, where 't' is a boolean array and 'a' is
        #            a float32, shouldn't promote to float64
        #
        # a = np.array([1.0, 1.5], dtype=np.float32)
        # t = np.array([True, False])
        # b = t*a
        # assert_equal(b, [1.0, 0.0])
        # assert_equal(b.dtype, np.dtype('f4'))
        # b = (1-t)*a
        # assert_equal(b, [0.0, 1.5])
        # assert_equal(b.dtype, np.dtype('f4'))
        #
        # Probably ~t (bitwise negation) is more proper to use here,
        # but this is arguably less intuitive to understand at a glance, and
        # would fail if 't' is actually an integer array instead of boolean:
        #
        # b = (~t)*a
        # assert_equal(b, [0.0, 1.5])
        # assert_equal(b.dtype, np.dtype('f4'))

    def test_result_type(self):
        self.check_promotion_cases(np.result_type)
        assert_(np.result_type(None) == np.dtype(None))

    def test_promote_types_endian(self):
        # promote_types should always return native-endian types
        assert_equal(np.promote_types('<i8', '<i8'), np.dtype('i8'))
        assert_equal(np.promote_types('>i8', '>i8'), np.dtype('i8'))

        assert_equal(np.promote_types('>i8', '>U16'), np.dtype('U21'))
        assert_equal(np.promote_types('<i8', '<U16'), np.dtype('U21'))
        assert_equal(np.promote_types('>U16', '>i8'), np.dtype('U21'))
        assert_equal(np.promote_types('<U16', '<i8'), np.dtype('U21'))

        assert_equal(np.promote_types('<S5', '<U8'), np.dtype('U8'))
        assert_equal(np.promote_types('>S5', '>U8'), np.dtype('U8'))
        assert_equal(np.promote_types('<U8', '<S5'), np.dtype('U8'))
        assert_equal(np.promote_types('>U8', '>S5'), np.dtype('U8'))
        assert_equal(np.promote_types('<U5', '<U8'), np.dtype('U8'))
        assert_equal(np.promote_types('>U8', '>U5'), np.dtype('U8'))

        assert_equal(np.promote_types('<M8', '<M8'), np.dtype('M8'))
        assert_equal(np.promote_types('>M8', '>M8'), np.dtype('M8'))
        assert_equal(np.promote_types('<m8', '<m8'), np.dtype('m8'))
        assert_equal(np.promote_types('>m8', '>m8'), np.dtype('m8'))

    def test_can_cast_and_promote_usertypes(self):
        # The rational type defines safe casting for signed integers,
        # boolean. Rational itself *does* cast safely to double.
        # (rational does not actually cast to all signed integers, e.g.
        # int64 can be both long and longlong and it registers only the first)
        valid_types = ["int8", "int16", "int32", "int64", "bool"]
        invalid_types = "BHILQP" + "FDG" + "mM" + "f" + "V"

        rational_dt = np.dtype(rational)
        for numpy_dtype in valid_types:
            numpy_dtype = np.dtype(numpy_dtype)
            assert np.can_cast(numpy_dtype, rational_dt)
            assert np.promote_types(numpy_dtype, rational_dt) is rational_dt

        for numpy_dtype in invalid_types:
            numpy_dtype = np.dtype(numpy_dtype)
            assert not np.can_cast(numpy_dtype, rational_dt)
            with pytest.raises(TypeError):
                np.promote_types(numpy_dtype, rational_dt)

        double_dt = np.dtype("double")
        assert np.can_cast(rational_dt, double_dt)
        assert np.promote_types(double_dt, rational_dt) is double_dt

    @pytest.mark.parametrize("swap", ["", "swap"])
    @pytest.mark.parametrize("string_dtype", ["U", "S"])
    def test_promote_types_strings(self, swap, string_dtype):
        if swap == "swap":
            promote_types = lambda a, b: np.promote_types(b, a)
        else:
            promote_types = np.promote_types

        S = string_dtype

        # Promote numeric with unsized string:
        assert_equal(promote_types('bool', S), np.dtype(S + '5'))
        assert_equal(promote_types('b', S), np.dtype(S + '4'))
        assert_equal(promote_types('u1', S), np.dtype(S + '3'))
        assert_equal(promote_types('u2', S), np.dtype(S + '5'))
        assert_equal(promote_types('u4', S), np.dtype(S + '10'))
        assert_equal(promote_types('u8', S), np.dtype(S + '20'))
        assert_equal(promote_types('i1', S), np.dtype(S + '4'))
        assert_equal(promote_types('i2', S), np.dtype(S + '6'))
        assert_equal(promote_types('i4', S), np.dtype(S + '11'))
        assert_equal(promote_types('i8', S), np.dtype(S + '21'))
        # Promote numeric with sized string:
        assert_equal(promote_types('bool', S + '1'), np.dtype(S + '5'))
        assert_equal(promote_types('bool', S + '30'), np.dtype(S + '30'))
        assert_equal(promote_types('b', S + '1'), np.dtype(S + '4'))
        assert_equal(promote_types('b', S + '30'), np.dtype(S + '30'))
        assert_equal(promote_types('u1', S + '1'), np.dtype(S + '3'))
        assert_equal(promote_types('u1', S + '30'), np.dtype(S + '30'))
        assert_equal(promote_types('u2', S + '1'), np.dtype(S + '5'))
        assert_equal(promote_types('u2', S + '30'), np.dtype(S + '30'))
        assert_equal(promote_types('u4', S + '1'), np.dtype(S + '10'))
        assert_equal(promote_types('u4', S + '30'), np.dtype(S + '30'))
        assert_equal(promote_types('u8', S + '1'), np.dtype(S + '20'))
        assert_equal(promote_types('u8', S + '30'), np.dtype(S + '30'))
        # Promote with object:
        assert_equal(promote_types('O', S + '30'), np.dtype('O'))

    @pytest.mark.parametrize(["dtype1", "dtype2"],
            [[np.dtype("V6"), np.dtype("V10")],  # mismatch shape
             # Mismatching names:
             [np.dtype([("name1", "i8")]), np.dtype([("name2", "i8")])],
            ])
    def test_invalid_void_promotion(self, dtype1, dtype2):
        with pytest.raises(TypeError):
            np.promote_types(dtype1, dtype2)

    @pytest.mark.parametrize(["dtype1", "dtype2"],
            [[np.dtype("V10"), np.dtype("V10")],
             [np.dtype([("name1", "i8")]),
              np.dtype([("name1", np.dtype("i8").newbyteorder())])],
             [np.dtype("i8,i8"), np.dtype("i8,>i8")],
             [np.dtype("i8,i8"), np.dtype("i4,i4")],
            ])
    def test_valid_void_promotion(self, dtype1, dtype2):
        assert np.promote_types(dtype1, dtype2) == dtype1

    @pytest.mark.parametrize("dtype",
            list(np.typecodes["All"]) +
            ["i,i", "10i", "S3", "S100", "U3", "U100", rational])
    def test_promote_identical_types_metadata(self, dtype):
        # The same type passed in twice to promote types always
        # preserves metadata
        metadata = {1: 1}
        dtype = np.dtype(dtype, metadata=metadata)

        res = np.promote_types(dtype, dtype)
        assert res.metadata == dtype.metadata

        # byte-swapping preserves and makes the dtype native:
        dtype = dtype.newbyteorder()
        if dtype.isnative:
            # The type does not have byte swapping
            return

        res = np.promote_types(dtype, dtype)

        # Metadata is (currently) generally lost on byte-swapping (except for
        # unicode.
        if dtype.char != "U":
            assert res.metadata is None
        else:
            assert res.metadata == metadata
        assert res.isnative

    @pytest.mark.slow
    @pytest.mark.filterwarnings('ignore:Promotion of numbers:FutureWarning')
    @pytest.mark.parametrize(["dtype1", "dtype2"],
            itertools.product(
                list(np.typecodes["All"]) +
                ["i,i", "S3", "S100", "U3", "U100", rational],
                repeat=2))
    def test_promote_types_metadata(self, dtype1, dtype2):
        """Metadata handling in promotion does not appear formalized
        right now in NumPy. This test should thus be considered to
        document behaviour, rather than test the correct definition of it.

        This test is very ugly, it was useful for rewriting part of the
        promotion, but probably should eventually be replaced/deleted
        (i.e. when metadata handling in promotion is better defined).
        """
        metadata1 = {1: 1}
        metadata2 = {2: 2}
        dtype1 = np.dtype(dtype1, metadata=metadata1)
        dtype2 = np.dtype(dtype2, metadata=metadata2)

        try:
            res = np.promote_types(dtype1, dtype2)
        except TypeError:
            # Promotion failed, this test only checks metadata
            return

        if res.char not in "USV" or res.names is not None or res.shape != ():
            # All except string dtypes (and unstructured void) lose metadata
            # on promotion (unless both dtypes are identical).
            # At some point structured ones did not, but were restrictive.
            assert res.metadata is None
        elif res == dtype1:
            # If one result is the result, it is usually returned unchanged:
            assert res is dtype1
        elif res == dtype2:
            # dtype1 may have been cast to the same type/kind as dtype2.
            # If the resulting dtype is identical we currently pick the cast
            # version of dtype1, which lost the metadata:
            if np.promote_types(dtype1, dtype2.kind) == dtype2:
                res.metadata is None
            else:
                res.metadata == metadata2
        else:
            assert res.metadata is None

        # Try again for byteswapped version
        dtype1 = dtype1.newbyteorder()
        assert dtype1.metadata == metadata1
        res_bs = np.promote_types(dtype1, dtype2)
        assert res_bs == res
        assert res_bs.metadata == res.metadata

    def test_can_cast(self):
        assert_(np.can_cast(np.int32, np.int64))
        assert_(np.can_cast(np.float64, complex))
        assert_(not np.can_cast(complex, float))

        assert_(np.can_cast('i8', 'f8'))
        assert_(not np.can_cast('i8', 'f4'))
        assert_(np.can_cast('i4', 'S11'))

        assert_(np.can_cast('i8', 'i8', 'no'))
        assert_(not np.can_cast('<i8', '>i8', 'no'))

        assert_(np.can_cast('<i8', '>i8', 'equiv'))
        assert_(not np.can_cast('<i4', '>i8', 'equiv'))

        assert_(np.can_cast('<i4', '>i8', 'safe'))
        assert_(not np.can_cast('<i8', '>i4', 'safe'))

        assert_(np.can_cast('<i8', '>i4', 'same_kind'))
        assert_(not np.can_cast('<i8', '>u4', 'same_kind'))

        assert_(np.can_cast('<i8', '>u4', 'unsafe'))

        assert_(np.can_cast('bool', 'S5'))
        assert_(not np.can_cast('bool', 'S4'))

        assert_(np.can_cast('b', 'S4'))
        assert_(not np.can_cast('b', 'S3'))

        assert_(np.can_cast('u1', 'S3'))
        assert_(not np.can_cast('u1', 'S2'))
        assert_(np.can_cast('u2', 'S5'))
        assert_(not np.can_cast('u2', 'S4'))
        assert_(np.can_cast('u4', 'S10'))
        assert_(not np.can_cast('u4', 'S9'))
        assert_(np.can_cast('u8', 'S20'))
        assert_(not np.can_cast('u8', 'S19'))

        assert_(np.can_cast('i1', 'S4'))
        assert_(not np.can_cast('i1', 'S3'))
        assert_(np.can_cast('i2', 'S6'))
        assert_(not np.can_cast('i2', 'S5'))
        assert_(np.can_cast('i4', 'S11'))
        assert_(not np.can_cast('i4', 'S10'))
        assert_(np.can_cast('i8', 'S21'))
        assert_(not np.can_cast('i8', 'S20'))

        assert_(np.can_cast('bool', 'S5'))
        assert_(not np.can_cast('bool', 'S4'))

        assert_(np.can_cast('b', 'U4'))
        assert_(not np.can_cast('b', 'U3'))

        assert_(np.can_cast('u1', 'U3'))
        assert_(not np.can_cast('u1', 'U2'))
        assert_(np.can_cast('u2', 'U5'))
        assert_(not np.can_cast('u2', 'U4'))
        assert_(np.can_cast('u4', 'U10'))
        assert_(not np.can_cast('u4', 'U9'))
        assert_(np.can_cast('u8', 'U20'))
        assert_(not np.can_cast('u8', 'U19'))

        assert_(np.can_cast('i1', 'U4'))
        assert_(not np.can_cast('i1', 'U3'))
        assert_(np.can_cast('i2', 'U6'))
        assert_(not np.can_cast('i2', 'U5'))
        assert_(np.can_cast('i4', 'U11'))
        assert_(not np.can_cast('i4', 'U10'))
        assert_(np.can_cast('i8', 'U21'))
        assert_(not np.can_cast('i8', 'U20'))

        assert_raises(TypeError, np.can_cast, 'i4', None)
        assert_raises(TypeError, np.can_cast, None, 'i4')

        # Also test keyword arguments
        assert_(np.can_cast(from_=np.int32, to=np.int64))

    def test_can_cast_simple_to_structured(self):
        # Non-structured can only be cast to structured in 'unsafe' mode.
        assert_(not np.can_cast('i4', 'i4,i4'))
        assert_(not np.can_cast('i4', 'i4,i2'))
        assert_(np.can_cast('i4', 'i4,i4', casting='unsafe'))
        assert_(np.can_cast('i4', 'i4,i2', casting='unsafe'))
        # Even if there is just a single field which is OK.
        assert_(not np.can_cast('i2', [('f1', 'i4')]))
        assert_(not np.can_cast('i2', [('f1', 'i4')], casting='same_kind'))
        assert_(np.can_cast('i2', [('f1', 'i4')], casting='unsafe'))
        # It should be the same for recursive structured or subarrays.
        assert_(not np.can_cast('i2', [('f1', 'i4,i4')]))
        assert_(np.can_cast('i2', [('f1', 'i4,i4')], casting='unsafe'))
        assert_(not np.can_cast('i2', [('f1', '(2,3)i4')]))
        assert_(np.can_cast('i2', [('f1', '(2,3)i4')], casting='unsafe'))

    def test_can_cast_structured_to_simple(self):
        # Need unsafe casting for structured to simple.
        assert_(not np.can_cast([('f1', 'i4')], 'i4'))
        assert_(np.can_cast([('f1', 'i4')], 'i4', casting='unsafe'))
        assert_(np.can_cast([('f1', 'i4')], 'i2', casting='unsafe'))
        # Since it is unclear what is being cast, multiple fields to
        # single should not work even for unsafe casting.
        assert_(not np.can_cast('i4,i4', 'i4', casting='unsafe'))
        # But a single field inside a single field is OK.
        assert_(not np.can_cast([('f1', [('x', 'i4')])], 'i4'))
        assert_(np.can_cast([('f1', [('x', 'i4')])], 'i4', casting='unsafe'))
        # And a subarray is fine too - it will just take the first element
        # (arguably not very consistently; might also take the first field).
        assert_(not np.can_cast([('f0', '(3,)i4')], 'i4'))
        assert_(np.can_cast([('f0', '(3,)i4')], 'i4', casting='unsafe'))
        # But a structured subarray with multiple fields should fail.
        assert_(not np.can_cast([('f0', ('i4,i4'), (2,))], 'i4',
                                casting='unsafe'))

    def test_can_cast_values(self):
        # With NumPy 2 and NEP 50, can_cast errors on Python scalars.  We could
        # define this as (usually safe) at some point, and already do so
        # in `copyto` and ufuncs (but there an error is raised if the integer
        # is out of bounds and a warning for out-of-bound floats).
        # Raises even for unsafe, previously checked within range (for floats
        # that was approximately whether it would overflow to inf).
        with pytest.raises(TypeError):
            np.can_cast(4, "int8", casting="unsafe")

        with pytest.raises(TypeError):
            np.can_cast(4.0, "float64", casting="unsafe")

        with pytest.raises(TypeError):
            np.can_cast(4j, "complex128", casting="unsafe")

    @pytest.mark.parametrize("dtype",
            list("?bhilqBHILQefdgFDG") + [rational])
    def test_can_cast_scalars(self, dtype):
        # Basic test to ensure that scalars are supported in can-cast
        # (does not check behavior exhaustively).
        dtype = np.dtype(dtype)
        scalar = dtype.type(0)

        assert np.can_cast(scalar, "int64") == np.can_cast(dtype, "int64")
        assert np.can_cast(scalar, "float32", casting="unsafe")


# Custom exception class to test exception propagation in fromiter
class NIterError(Exception):
    pass


class TestFromiter:
    def makegen(self):
        return (x**2 for x in range(24))

    def test_types(self):
        ai32 = np.fromiter(self.makegen(), np.int32)
        ai64 = np.fromiter(self.makegen(), np.int64)
        af = np.fromiter(self.makegen(), float)
        assert_(ai32.dtype == np.dtype(np.int32))
        assert_(ai64.dtype == np.dtype(np.int64))
        assert_(af.dtype == np.dtype(float))

    def test_lengths(self):
        expected = np.array(list(self.makegen()))
        a = np.fromiter(self.makegen(), int)
        a20 = np.fromiter(self.makegen(), int, 20)
        assert_(len(a) == len(expected))
        assert_(len(a20) == 20)
        assert_raises(ValueError, np.fromiter,
                          self.makegen(), int, len(expected) + 10)

    def test_values(self):
        expected = np.array(list(self.makegen()))
        a = np.fromiter(self.makegen(), int)
        a20 = np.fromiter(self.makegen(), int, 20)
        assert_(np.all(a == expected, axis=0))
        assert_(np.all(a20 == expected[:20], axis=0))

    def load_data(self, n, eindex):
        # Utility method for the issue 2592 tests.
        # Raise an exception at the desired index in the iterator.
        for e in range(n):
            if e == eindex:
                raise NIterError(f'error at index {eindex}')
            yield e

    @pytest.mark.parametrize("dtype", [int, object])
    @pytest.mark.parametrize(["count", "error_index"], [(10, 5), (10, 9)])
    def test_2592(self, count, error_index, dtype):
        # Test iteration exceptions are correctly raised. The data/generator
        # has `count` elements but errors at `error_index`
        iterable = self.load_data(count, error_index)
        with pytest.raises(NIterError):
            np.fromiter(iterable, dtype=dtype, count=count)

    @pytest.mark.parametrize("dtype", ["S", "S0", "V0", "U0"])
    def test_empty_not_structured(self, dtype):
        # Note, "S0" could be allowed at some point, so long "S" (without
        # any length) is rejected.
        with pytest.raises(ValueError, match="Must specify length"):
            np.fromiter([], dtype=dtype)

    @pytest.mark.parametrize(["dtype", "data"],
            [("d", [1, 2, 3, 4, 5, 6, 7, 8, 9]),
             ("O", [1, 2, 3, 4, 5, 6, 7, 8, 9]),
             ("i,O", [(1, 2), (5, 4), (2, 3), (9, 8), (6, 7)]),
             # subarray dtypes (important because their dimensions end up
             # in the result arrays dimension:
             ("2i", [(1, 2), (5, 4), (2, 3), (9, 8), (6, 7)]),
             (np.dtype(("O", (2, 3))),
              [((1, 2, 3), (3, 4, 5)), ((3, 2, 1), (5, 4, 3))])])
    @pytest.mark.parametrize("length_hint", [0, 1])
    def test_growth_and_complicated_dtypes(self, dtype, data, length_hint):
        dtype = np.dtype(dtype)

        data = data * 100  # make sure we realloc a bit

        class MyIter:
            # Class/example from gh-15789
            def __length_hint__(self):
                # only required to be an estimate, this is legal
                return length_hint  # 0 or 1

            def __iter__(self):
                return iter(data)

        res = np.fromiter(MyIter(), dtype=dtype)
        expected = np.array(data, dtype=dtype)

        assert_array_equal(res, expected)

    def test_empty_result(self):
        class MyIter:
            def __length_hint__(self):
                return 10

            def __iter__(self):
                return iter([])  # actual iterator is empty.

        res = np.fromiter(MyIter(), dtype="d")
        assert res.shape == (0,)
        assert res.dtype == "d"

    def test_too_few_items(self):
        msg = "iterator too short: Expected 10 but iterator had only 3 items."
        with pytest.raises(ValueError, match=msg):
            np.fromiter([1, 2, 3], count=10, dtype=int)

    def test_failed_itemsetting(self):
        with pytest.raises(TypeError):
            np.fromiter([1, None, 3], dtype=int)

        # The following manages to hit somewhat trickier code paths:
        iterable = ((2, 3, 4) for i in range(5))
        with pytest.raises(ValueError):
            np.fromiter(iterable, dtype=np.dtype((int, 2)))

class TestNonzero:
    def test_nonzero_trivial(self):
        assert_equal(np.count_nonzero(np.array([])), 0)
        assert_equal(np.count_nonzero(np.array([], dtype='?')), 0)
        assert_equal(np.nonzero(np.array([])), ([],))

        assert_equal(np.count_nonzero(np.array([0])), 0)
        assert_equal(np.count_nonzero(np.array([0], dtype='?')), 0)
        assert_equal(np.nonzero(np.array([0])), ([],))

        assert_equal(np.count_nonzero(np.array([1])), 1)
        assert_equal(np.count_nonzero(np.array([1], dtype='?')), 1)
        assert_equal(np.nonzero(np.array([1])), ([0],))

    def test_nonzero_zerodim(self):
        err_msg = "Calling nonzero on 0d arrays is not allowed"
        with assert_raises_regex(ValueError, err_msg):
            np.nonzero(np.array(0))
        with assert_raises_regex(ValueError, err_msg):
            np.array(1).nonzero()

    def test_nonzero_onedim(self):
        x = np.array([1, 0, 2, -1, 0, 0, 8])
        assert_equal(np.count_nonzero(x), 4)
        assert_equal(np.count_nonzero(x), 4)
        assert_equal(np.nonzero(x), ([0, 2, 3, 6],))

        # x = np.array([(1, 2), (0, 0), (1, 1), (-1, 3), (0, 7)],
        #              dtype=[('a', 'i4'), ('b', 'i2')])
        x = np.array([(1, 2, -5, -3), (0, 0, 2, 7), (1, 1, 0, 1), (-1, 3, 1, 0), (0, 7, 0, 4)],
                     dtype=[('a', 'i4'), ('b', 'i2'), ('c', 'i1'), ('d', 'i8')])
        assert_equal(np.count_nonzero(x['a']), 3)
        assert_equal(np.count_nonzero(x['b']), 4)
        assert_equal(np.count_nonzero(x['c']), 3)
        assert_equal(np.count_nonzero(x['d']), 4)
        assert_equal(np.nonzero(x['a']), ([0, 2, 3],))
        assert_equal(np.nonzero(x['b']), ([0, 2, 3, 4],))

    def test_nonzero_twodim(self):
        x = np.array([[0, 1, 0], [2, 0, 3]])
        assert_equal(np.count_nonzero(x.astype('i1')), 3)
        assert_equal(np.count_nonzero(x.astype('i2')), 3)
        assert_equal(np.count_nonzero(x.astype('i4')), 3)
        assert_equal(np.count_nonzero(x.astype('i8')), 3)
        assert_equal(np.nonzero(x), ([0, 1, 1], [1, 0, 2]))

        x = np.eye(3)
        assert_equal(np.count_nonzero(x.astype('i1')), 3)
        assert_equal(np.count_nonzero(x.astype('i2')), 3)
        assert_equal(np.count_nonzero(x.astype('i4')), 3)
        assert_equal(np.count_nonzero(x.astype('i8')), 3)
        assert_equal(np.nonzero(x), ([0, 1, 2], [0, 1, 2]))

        x = np.array([[(0, 1), (0, 0), (1, 11)],
                   [(1, 1), (1, 0), (0, 0)],
                   [(0, 0), (1, 5), (0, 1)]], dtype=[('a', 'f4'), ('b', 'u1')])
        assert_equal(np.count_nonzero(x['a']), 4)
        assert_equal(np.count_nonzero(x['b']), 5)
        assert_equal(np.nonzero(x['a']), ([0, 1, 1, 2], [2, 0, 1, 1]))
        assert_equal(np.nonzero(x['b']), ([0, 0, 1, 2, 2], [0, 2, 0, 1, 2]))

        assert_(not x['a'].T.flags.aligned)
        assert_equal(np.count_nonzero(x['a'].T), 4)
        assert_equal(np.count_nonzero(x['b'].T), 5)
        assert_equal(np.nonzero(x['a'].T), ([0, 1, 1, 2], [1, 1, 2, 0]))
        assert_equal(np.nonzero(x['b'].T), ([0, 0, 1, 2, 2], [0, 1, 2, 0, 2]))

    def test_sparse(self):
        # test special sparse condition boolean code path
        for i in range(20):
            c = np.zeros(200, dtype=bool)
            c[i::20] = True
            assert_equal(np.nonzero(c)[0], np.arange(i, 200 + i, 20))

            c = np.zeros(400, dtype=bool)
            c[10 + i:20 + i] = True
            c[20 + i * 2] = True
            assert_equal(np.nonzero(c)[0],
                         np.concatenate((np.arange(10 + i, 20 + i), [20 + i * 2])))

    @pytest.mark.parametrize('dtype', [np.float32, np.float64])
    def test_nonzero_float_dtypes(self, dtype):
        rng = np.random.default_rng(seed=10)
        x = ((2**33) * rng.normal(size=100)).astype(dtype)
        x[rng.choice(50, size=100)] = 0
        idxs = np.nonzero(x)[0]
        assert_equal(np.array_equal(np.where(x != 0)[0], idxs), True)

    @pytest.mark.parametrize('dtype', [bool, np.int8, np.int16, np.int32, np.int64,
                                       np.uint8, np.uint16, np.uint32, np.uint64])
    def test_nonzero_integer_dtypes(self, dtype):
        rng = np.random.default_rng(seed=10)
        x = rng.integers(0, 255, size=100).astype(dtype)
        x[rng.choice(50, size=100)] = 0
        idxs = np.nonzero(x)[0]
        assert_equal(np.array_equal(np.where(x != 0)[0], idxs), True)

    def test_return_type(self):
        class C(np.ndarray):
            pass

        for view in (C, np.ndarray):
            for nd in range(1, 4):
                shape = tuple(range(2, 2 + nd))
                x = np.arange(np.prod(shape)).reshape(shape).view(view)
                for nzx in (np.nonzero(x), x.nonzero()):
                    for nzx_i in nzx:
                        assert_(type(nzx_i) is np.ndarray)
                        assert_(nzx_i.flags.writeable)

    def test_count_nonzero_axis(self):
        # Basic check of functionality
        m = np.array([[0, 1, 7, 0, 0], [3, 0, 0, 2, 19]])

        expected = np.array([1, 1, 1, 1, 1])
        assert_equal(np.count_nonzero(m, axis=0), expected)

        expected = np.array([2, 3])
        assert_equal(np.count_nonzero(m, axis=1), expected)

        assert_raises(ValueError, np.count_nonzero, m, axis=(1, 1))
        assert_raises(TypeError, np.count_nonzero, m, axis='foo')
        assert_raises(AxisError, np.count_nonzero, m, axis=3)
        assert_raises(TypeError, np.count_nonzero,
                      m, axis=np.array([[1], [2]]))

    def test_count_nonzero_axis_all_dtypes(self):
        # More thorough test that the axis argument is respected
        # for all dtypes and responds correctly when presented with
        # either integer or tuple arguments for axis
        msg = "Mismatch for dtype: %s"

        def assert_equal_w_dt(a, b, err_msg):
            assert_equal(a.dtype, b.dtype, err_msg=err_msg)
            assert_equal(a, b, err_msg=err_msg)

        for dt in np.typecodes['All']:
            err_msg = msg % (np.dtype(dt).name,)

            if dt != 'V':
                if dt != 'M':
                    m = np.zeros((3, 3), dtype=dt)
                    n = np.ones(1, dtype=dt)

                    m[0, 0] = n[0]
                    m[1, 0] = n[0]

                else:  # np.zeros doesn't work for np.datetime64
                    m = np.array(['1970-01-01'] * 9)
                    m = m.reshape((3, 3))

                    m[0, 0] = '1970-01-12'
                    m[1, 0] = '1970-01-12'
                    m = m.astype(dt)

                expected = np.array([2, 0, 0], dtype=np.intp)
                assert_equal_w_dt(np.count_nonzero(m, axis=0),
                                  expected, err_msg=err_msg)

                expected = np.array([1, 1, 0], dtype=np.intp)
                assert_equal_w_dt(np.count_nonzero(m, axis=1),
                                  expected, err_msg=err_msg)

                expected = np.array(2)
                assert_equal(np.count_nonzero(m, axis=(0, 1)),
                             expected, err_msg=err_msg)
                assert_equal(np.count_nonzero(m, axis=None),
                             expected, err_msg=err_msg)
                assert_equal(np.count_nonzero(m),
                             expected, err_msg=err_msg)

            if dt == 'V':
                # There are no 'nonzero' objects for np.void, so the testing
                # setup is slightly different for this dtype
                m = np.array([np.void(1)] * 6).reshape((2, 3))

                expected = np.array([0, 0, 0], dtype=np.intp)
                assert_equal_w_dt(np.count_nonzero(m, axis=0),
                                  expected, err_msg=err_msg)

                expected = np.array([0, 0], dtype=np.intp)
                assert_equal_w_dt(np.count_nonzero(m, axis=1),
                                  expected, err_msg=err_msg)

                expected = np.array(0)
                assert_equal(np.count_nonzero(m, axis=(0, 1)),
                             expected, err_msg=err_msg)
                assert_equal(np.count_nonzero(m, axis=None),
                             expected, err_msg=err_msg)
                assert_equal(np.count_nonzero(m),
                             expected, err_msg=err_msg)

    def test_count_nonzero_axis_consistent(self):
        # Check that the axis behaviour for valid axes in
        # non-special cases is consistent (and therefore
        # correct) by checking it against an integer array
        # that is then casted to the generic object dtype
        from itertools import combinations, permutations

        axis = (0, 1, 2, 3)
        size = (5, 5, 5, 5)
        msg = "Mismatch for axis: %s"

        rng = np.random.RandomState(1234)
        m = rng.randint(-100, 100, size=size)
        n = m.astype(object)

        for length in range(len(axis)):
            for combo in combinations(axis, length):
                for perm in permutations(combo):
                    assert_equal(
                        np.count_nonzero(m, axis=perm),
                        np.count_nonzero(n, axis=perm),
                        err_msg=msg % (perm,))

    def test_countnonzero_axis_empty(self):
        a = np.array([[0, 0, 1], [1, 0, 1]])
        assert_equal(np.count_nonzero(a, axis=()), a.astype(bool))

    def test_countnonzero_keepdims(self):
        a = np.array([[0, 0, 1, 0],
                      [0, 3, 5, 0],
                      [7, 9, 2, 0]])
        assert_equal(np.count_nonzero(a, axis=0, keepdims=True),
                     [[1, 2, 3, 0]])
        assert_equal(np.count_nonzero(a, axis=1, keepdims=True),
                     [[1], [2], [3]])
        assert_equal(np.count_nonzero(a, keepdims=True),
                     [[6]])

    def test_array_method(self):
        # Tests that the array method
        # call to nonzero works
        m = np.array([[1, 0, 0], [4, 0, 6]])
        tgt = [[0, 1, 1], [0, 0, 2]]

        assert_equal(m.nonzero(), tgt)

    def test_nonzero_invalid_object(self):
        # gh-9295
        a = np.array([np.array([1, 2]), 3], dtype=object)
        assert_raises(ValueError, np.nonzero, a)

        class BoolErrors:
            def __bool__(self):
                raise ValueError("Not allowed")

        assert_raises(ValueError, np.nonzero, np.array([BoolErrors()]))

    def test_nonzero_sideeffect_safety(self):
        # gh-13631
        class FalseThenTrue:
            _val = False

            def __bool__(self):
                try:
                    return self._val
                finally:
                    self._val = True

        class TrueThenFalse:
            _val = True

            def __bool__(self):
                try:
                    return self._val
                finally:
                    self._val = False

        # result grows on the second pass
        a = np.array([True, FalseThenTrue()])
        assert_raises(RuntimeError, np.nonzero, a)

        a = np.array([[True], [FalseThenTrue()]])
        assert_raises(RuntimeError, np.nonzero, a)

        # result shrinks on the second pass
        a = np.array([False, TrueThenFalse()])
        assert_raises(RuntimeError, np.nonzero, a)

        a = np.array([[False], [TrueThenFalse()]])
        assert_raises(RuntimeError, np.nonzero, a)

    def test_nonzero_sideffects_structured_void(self):
        # Checks that structured void does not mutate alignment flag of
        # original array.
        arr = np.zeros(5, dtype="i1,i8,i8")  # `ones` may short-circuit
        assert arr.flags.aligned  # structs are considered "aligned"
        assert not arr["f2"].flags.aligned
        # make sure that nonzero/count_nonzero do not flip the flag:
        np.nonzero(arr)
        assert arr.flags.aligned
        np.count_nonzero(arr)
        assert arr.flags.aligned

    def test_nonzero_exception_safe(self):
        # gh-13930

        class ThrowsAfter:
            def __init__(self, iters):
                self.iters_left = iters

            def __bool__(self):
                if self.iters_left == 0:
                    raise ValueError("called `iters` times")

                self.iters_left -= 1
                return True

        """
        Test that a ValueError is raised instead of a SystemError

        If the __bool__ function is called after the error state is set,
        Python (cpython) will raise a SystemError.
        """

        # assert that an exception in first pass is handled correctly
        a = np.array([ThrowsAfter(5)] * 10)
        assert_raises(ValueError, np.nonzero, a)

        # raise exception in second pass for 1-dimensional loop
        a = np.array([ThrowsAfter(15)] * 10)
        assert_raises(ValueError, np.nonzero, a)

        # raise exception in second pass for n-dimensional loop
        a = np.array([[ThrowsAfter(15)]] * 10)
        assert_raises(ValueError, np.nonzero, a)

    def test_nonzero_byteorder(self):
        values = [0., -0., 1, float('nan'), 0, 1,
                  np.float16(0), np.float16(12.3)]
        expected_values = [0, 0, 1, 1, 0, 1, 0, 1]

        for value, expected in zip(values, expected_values):
            A = np.array([value])
            A_byteswapped = (A.view(A.dtype.newbyteorder()).byteswap()).copy()

            assert np.count_nonzero(A) == expected
            assert np.count_nonzero(A_byteswapped) == expected

    def test_count_nonzero_non_aligned_array(self):
        # gh-27523
        b = np.zeros(64 + 1, dtype=np.int8)[1:]
        b = b.view(int)
        b[:] = np.arange(b.size)
        b[::2] = 0
        assert b.flags.aligned is False
        assert np.count_nonzero(b) == b.size / 2

        b = np.zeros(64 + 1, dtype=np.float16)[1:]
        b = b.view(float)
        b[:] = np.arange(b.size)
        b[::2] = 0
        assert b.flags.aligned is False
        assert np.count_nonzero(b) == b.size / 2


class TestIndex:
    def test_boolean(self):
        a = rand(3, 5, 8)
        V = rand(5, 8)
        g1 = randint(0, 5, size=15)
        g2 = randint(0, 8, size=15)
        V[g1, g2] = -V[g1, g2]
        assert_((np.array([a[0][V > 0], a[1][V > 0], a[2][V > 0]]) == a[:, V > 0]).all())

    def test_boolean_edgecase(self):
        a = np.array([], dtype='int32')
        b = np.array([], dtype='bool')
        c = a[b]
        assert_equal(c, [])
        assert_equal(c.dtype, np.dtype('int32'))


class TestBinaryRepr:
    def test_zero(self):
        assert_equal(np.binary_repr(0), '0')

    def test_positive(self):
        assert_equal(np.binary_repr(10), '1010')
        assert_equal(np.binary_repr(12522),
                     '11000011101010')
        assert_equal(np.binary_repr(10736848),
                     '101000111101010011010000')

    def test_negative(self):
        assert_equal(np.binary_repr(-1), '-1')
        assert_equal(np.binary_repr(-10), '-1010')
        assert_equal(np.binary_repr(-12522),
                     '-11000011101010')
        assert_equal(np.binary_repr(-10736848),
                     '-101000111101010011010000')

    def test_sufficient_width(self):
        assert_equal(np.binary_repr(0, width=5), '00000')
        assert_equal(np.binary_repr(10, width=7), '0001010')
        assert_equal(np.binary_repr(-5, width=7), '1111011')

    def test_neg_width_boundaries(self):
        # see gh-8670

        # Ensure that the example in the issue does not
        # break before proceeding to a more thorough test.
        assert_equal(np.binary_repr(-128, width=8), '10000000')

        for width in range(1, 11):
            num = -2**(width - 1)
            exp = '1' + (width - 1) * '0'
            assert_equal(np.binary_repr(num, width=width), exp)

    def test_large_neg_int64(self):
        # See gh-14289.
        assert_equal(np.binary_repr(np.int64(-2**62), width=64),
                     '11' + '0' * 62)


class TestBaseRepr:
    def test_base3(self):
        assert_equal(np.base_repr(3**5, 3), '100000')

    def test_positive(self):
        assert_equal(np.base_repr(12, 10), '12')
        assert_equal(np.base_repr(12, 10, 4), '000012')
        assert_equal(np.base_repr(12, 4), '30')
        assert_equal(np.base_repr(3731624803700888, 36), '10QR0ROFCEW')

    def test_negative(self):
        assert_equal(np.base_repr(-12, 10), '-12')
        assert_equal(np.base_repr(-12, 10, 4), '-000012')
        assert_equal(np.base_repr(-12, 4), '-30')

    def test_base_range(self):
        with assert_raises(ValueError):
            np.base_repr(1, 1)
        with assert_raises(ValueError):
            np.base_repr(1, 37)

    def test_minimal_signed_int(self):
        assert_equal(np.base_repr(np.int8(-128)), '-10000000')


def _test_array_equal_parametrizations():
    """
    we pre-create arrays as we sometime want to pass the same instance
    and sometime not. Passing the same instances may not mean the array are
    equal, especially when containing None
    """
    # those are 0-d arrays, it used to be a special case
    # where (e0 == e0).all() would raise
    e0 = np.array(0, dtype="int")
    e1 = np.array(1, dtype="float")
    # x,y, nan_equal, expected_result
    yield (e0, e0.copy(), None, True)
    yield (e0, e0.copy(), False, True)
    yield (e0, e0.copy(), True, True)

    #
    yield (e1, e1.copy(), None, True)
    yield (e1, e1.copy(), False, True)
    yield (e1, e1.copy(), True, True)

    # Non-nanable - those cannot hold nans
    a12 = np.array([1, 2])
    a12b = a12.copy()
    a123 = np.array([1, 2, 3])
    a13 = np.array([1, 3])
    a34 = np.array([3, 4])

    aS1 = np.array(["a"], dtype="S1")
    aS1b = aS1.copy()
    aS1u4 = np.array([("a", 1)], dtype="S1,u4")
    aS1u4b = aS1u4.copy()

    yield (a12, a12b, None, True)
    yield (a12, a12, None, True)
    yield (a12, a123, None, False)
    yield (a12, a34, None, False)
    yield (a12, a13, None, False)
    yield (aS1, aS1b, None, True)
    yield (aS1, aS1, None, True)

    # Non-float dtype - equal_nan should have no effect,
    yield (a123, a123, None, True)
    yield (a123, a123, False, True)
    yield (a123, a123, True, True)
    yield (a123, a123.copy(), None, True)
    yield (a123, a123.copy(), False, True)
    yield (a123, a123.copy(), True, True)
    yield (a123.astype("float"), a123.astype("float"), None, True)
    yield (a123.astype("float"), a123.astype("float"), False, True)
    yield (a123.astype("float"), a123.astype("float"), True, True)

    # these can hold None
    b1 = np.array([1, 2, np.nan])
    b2 = np.array([1, np.nan, 2])
    b3 = np.array([1, 2, np.inf])
    b4 = np.array(np.nan)

    # instances are the same
    yield (b1, b1, None, False)
    yield (b1, b1, False, False)
    yield (b1, b1, True, True)

    # equal but not same instance
    yield (b1, b1.copy(), None, False)
    yield (b1, b1.copy(), False, False)
    yield (b1, b1.copy(), True, True)

    # same once stripped of Nan
    yield (b1, b2, None, False)
    yield (b1, b2, False, False)
    yield (b1, b2, True, False)

    # nan's not conflated with inf's
    yield (b1, b3, None, False)
    yield (b1, b3, False, False)
    yield (b1, b3, True, False)

    # all Nan
    yield (b4, b4, None, False)
    yield (b4, b4, False, False)
    yield (b4, b4, True, True)
    yield (b4, b4.copy(), None, False)
    yield (b4, b4.copy(), False, False)
    yield (b4, b4.copy(), True, True)

    t1 = b1.astype("timedelta64")
    t2 = b2.astype("timedelta64")

    # Timedeltas are particular
    yield (t1, t1, None, False)
    yield (t1, t1, False, False)
    yield (t1, t1, True, True)

    yield (t1, t1.copy(), None, False)
    yield (t1, t1.copy(), False, False)
    yield (t1, t1.copy(), True, True)

    yield (t1, t2, None, False)
    yield (t1, t2, False, False)
    yield (t1, t2, True, False)

    # Multi-dimensional array
    md1 = np.array([[0, 1], [np.nan, 1]])

    yield (md1, md1, None, False)
    yield (md1, md1, False, False)
    yield (md1, md1, True, True)
    yield (md1, md1.copy(), None, False)
    yield (md1, md1.copy(), False, False)
    yield (md1, md1.copy(), True, True)
    # both complexes are nan+nan.j but the same instance
    cplx1, cplx2 = [np.array([np.nan + np.nan * 1j])] * 2

    # only real or img are nan.
    cplx3, cplx4 = np.complex64(1, np.nan), np.complex64(np.nan, 1)

    # Complex values
    yield (cplx1, cplx2, None, False)
    yield (cplx1, cplx2, False, False)
    yield (cplx1, cplx2, True, True)

    # Complex values, 1+nan, nan+1j
    yield (cplx3, cplx4, None, False)
    yield (cplx3, cplx4, False, False)
    yield (cplx3, cplx4, True, True)


class TestArrayComparisons:
    @pytest.mark.parametrize(
        "bx,by,equal_nan,expected", _test_array_equal_parametrizations()
    )
    def test_array_equal_equal_nan(self, bx, by, equal_nan, expected):
        """
        This test array_equal for a few combinations:

        - are the two inputs the same object or not (same object may not
          be equal if contains NaNs)
        - Whether we should consider or not, NaNs, being equal.

        """
        if equal_nan is None:
            res = np.array_equal(bx, by)
        else:
            res = np.array_equal(bx, by, equal_nan=equal_nan)
        assert_(res is expected)
        assert_(type(res) is bool)

    def test_array_equal_different_scalar_types(self):
        # https://github.com/numpy/numpy/issues/27271
        a = np.array("foo")
        b = np.array(1)
        assert not np.array_equal(a, b)
        assert not np.array_equiv(a, b)

    def test_none_compares_elementwise(self):
        a = np.array([None, 1, None], dtype=object)
        assert_equal(a == None, [True, False, True])  # noqa: E711
        assert_equal(a != None, [False, True, False])  # noqa: E711

        a = np.ones(3)
        assert_equal(a == None, [False, False, False])  # noqa: E711
        assert_equal(a != None, [True, True, True])  # noqa: E711

    def test_array_equiv(self):
        res = np.array_equiv(np.array([1, 2]), np.array([1, 2]))
        assert_(res)
        assert_(type(res) is bool)
        res = np.array_equiv(np.array([1, 2]), np.array([1, 2, 3]))
        assert_(not res)
        assert_(type(res) is bool)
        res = np.array_equiv(np.array([1, 2]), np.array([3, 4]))
        assert_(not res)
        assert_(type(res) is bool)
        res = np.array_equiv(np.array([1, 2]), np.array([1, 3]))
        assert_(not res)
        assert_(type(res) is bool)

        res = np.array_equiv(np.array([1, 1]), np.array([1]))
        assert_(res)
        assert_(type(res) is bool)
        res = np.array_equiv(np.array([1, 1]), np.array([[1], [1]]))
        assert_(res)
        assert_(type(res) is bool)
        res = np.array_equiv(np.array([1, 2]), np.array([2]))
        assert_(not res)
        assert_(type(res) is bool)
        res = np.array_equiv(np.array([1, 2]), np.array([[1], [2]]))
        assert_(not res)
        assert_(type(res) is bool)
        res = np.array_equiv(np.array([1, 2]), np.array([[1, 2, 3], [4, 5, 6], [7, 8, 9]]))
        assert_(not res)
        assert_(type(res) is bool)

    @pytest.mark.parametrize("dtype", ["V0", "V3", "V10"])
    def test_compare_unstructured_voids(self, dtype):
        zeros = np.zeros(3, dtype=dtype)

        assert_array_equal(zeros, zeros)
        assert not (zeros != zeros).any()

        if dtype == "V0":
            # Can't test != of actually different data
            return

        nonzeros = np.array([b"1", b"2", b"3"], dtype=dtype)

        assert not (zeros == nonzeros).any()
        assert (zeros != nonzeros).all()


def assert_array_strict_equal(x, y):
    assert_array_equal(x, y)
    # Check flags, 32 bit arches typically don't provide 16 byte alignment
    if ((x.dtype.alignment <= 8 or
            np.intp().dtype.itemsize != 4) and
            sys.platform != 'win32'):
        assert_(x.flags == y.flags)
    else:
        assert_(x.flags.owndata == y.flags.owndata)
        assert_(x.flags.writeable == y.flags.writeable)
        assert_(x.flags.c_contiguous == y.flags.c_contiguous)
        assert_(x.flags.f_contiguous == y.flags.f_contiguous)
        assert_(x.flags.writebackifcopy == y.flags.writebackifcopy)
    # check endianness
    assert_(x.dtype.isnative == y.dtype.isnative)


class TestClip:
    def setup_method(self):
        self.nr = 5
        self.nc = 3

    def fastclip(self, a, m, M, out=None, **kwargs):
        return a.clip(m, M, out=out, **kwargs)

    def clip(self, a, m, M, out=None):
        # use a.choose to verify fastclip result
        selector = np.less(a, m) + 2 * np.greater(a, M)
        return selector.choose((a, m, M), out=out)

    # Handy functions
    def _generate_data(self, n, m):
        return randn(n, m)

    def _generate_data_complex(self, n, m):
        return randn(n, m) + 1.j * rand(n, m)

    def _generate_flt_data(self, n, m):
        return (randn(n, m)).astype(np.float32)

    def _neg_byteorder(self, a):
        a = np.asarray(a)
        if sys.byteorder == 'little':
            a = a.astype(a.dtype.newbyteorder('>'))
        else:
            a = a.astype(a.dtype.newbyteorder('<'))
        return a

    def _generate_non_native_data(self, n, m):
        data = randn(n, m)
        data = self._neg_byteorder(data)
        assert_(not data.dtype.isnative)
        return data

    def _generate_int_data(self, n, m):
        return (10 * rand(n, m)).astype(np.int64)

    def _generate_int32_data(self, n, m):
        return (10 * rand(n, m)).astype(np.int32)

    # Now the real test cases

    @pytest.mark.parametrize("dtype", '?bhilqpBHILQPefdgFDGO')
    def test_ones_pathological(self, dtype):
        # for preservation of behavior described in
        # gh-12519; amin > amax behavior may still change
        # in the future
        arr = np.ones(10, dtype=dtype)
        expected = np.zeros(10, dtype=dtype)
        actual = np.clip(arr, 1, 0)
        if dtype == 'O':
            assert actual.tolist() == expected.tolist()
        else:
            assert_equal(actual, expected)

    def test_simple_double(self):
        # Test native double input with scalar min/max.
        a = self._generate_data(self.nr, self.nc)
        m = 0.1
        M = 0.6
        ac = self.fastclip(a, m, M)
        act = self.clip(a, m, M)
        assert_array_strict_equal(ac, act)

    def test_simple_int(self):
        # Test native int input with scalar min/max.
        a = self._generate_int_data(self.nr, self.nc)
        a = a.astype(int)
        m = -2
        M = 4
        ac = self.fastclip(a, m, M)
        act = self.clip(a, m, M)
        assert_array_strict_equal(ac, act)

    def test_array_double(self):
        # Test native double input with array min/max.
        a = self._generate_data(self.nr, self.nc)
        m = np.zeros(a.shape)
        M = m + 0.5
        ac = self.fastclip(a, m, M)
        act = self.clip(a, m, M)
        assert_array_strict_equal(ac, act)

    def test_simple_nonnative(self):
        # Test non native double input with scalar min/max.
        # Test native double input with non native double scalar min/max.
        a = self._generate_non_native_data(self.nr, self.nc)
        m = -0.5
        M = 0.6
        ac = self.fastclip(a, m, M)
        act = self.clip(a, m, M)
        assert_array_equal(ac, act)

        # Test native double input with non native double scalar min/max.
        a = self._generate_data(self.nr, self.nc)
        m = -0.5
        M = self._neg_byteorder(0.6)
        assert_(not M.dtype.isnative)
        ac = self.fastclip(a, m, M)
        act = self.clip(a, m, M)
        assert_array_equal(ac, act)

    def test_simple_complex(self):
        # Test native complex input with native double scalar min/max.
        # Test native input with complex double scalar min/max.
        a = 3 * self._generate_data_complex(self.nr, self.nc)
        m = -0.5
        M = 1.
        ac = self.fastclip(a, m, M)
        act = self.clip(a, m, M)
        assert_array_strict_equal(ac, act)

        # Test native input with complex double scalar min/max.
        a = 3 * self._generate_data(self.nr, self.nc)
        m = -0.5 + 1.j
        M = 1. + 2.j
        ac = self.fastclip(a, m, M)
        act = self.clip(a, m, M)
        assert_array_strict_equal(ac, act)

    def test_clip_complex(self):
        # Address Issue gh-5354 for clipping complex arrays
        # Test native complex input without explicit min/max
        # ie, either min=None or max=None
        a = np.ones(10, dtype=complex)
        m = a.min()
        M = a.max()
        am = self.fastclip(a, m, None)
        aM = self.fastclip(a, None, M)
        assert_array_strict_equal(am, a)
        assert_array_strict_equal(aM, a)

    def test_clip_non_contig(self):
        # Test clip for non contiguous native input and native scalar min/max.
        a = self._generate_data(self.nr * 2, self.nc * 3)
        a = a[::2, ::3]
        assert_(not a.flags['F_CONTIGUOUS'])
        assert_(not a.flags['C_CONTIGUOUS'])
        ac = self.fastclip(a, -1.6, 1.7)
        act = self.clip(a, -1.6, 1.7)
        assert_array_strict_equal(ac, act)

    def test_simple_out(self):
        # Test native double input with scalar min/max.
        a = self._generate_data(self.nr, self.nc)
        m = -0.5
        M = 0.6
        ac = np.zeros(a.shape)
        act = np.zeros(a.shape)
        self.fastclip(a, m, M, ac)
        self.clip(a, m, M, act)
        assert_array_strict_equal(ac, act)

    @pytest.mark.parametrize("casting", [None, "unsafe"])
    def test_simple_int32_inout(self, casting):
        # Test native int32 input with double min/max and int32 out.
        a = self._generate_int32_data(self.nr, self.nc)
        m = np.float64(0)
        M = np.float64(2)
        ac = np.zeros(a.shape, dtype=np.int32)
        act = ac.copy()
        if casting is None:
            with pytest.raises(TypeError):
                self.fastclip(a, m, M, ac, casting=casting)
        else:
            # explicitly passing "unsafe" will silence warning
            self.fastclip(a, m, M, ac, casting=casting)
            self.clip(a, m, M, act)
            assert_array_strict_equal(ac, act)

    def test_simple_int64_out(self):
        # Test native int32 input with int32 scalar min/max and int64 out.
        a = self._generate_int32_data(self.nr, self.nc)
        m = np.int32(-1)
        M = np.int32(1)
        ac = np.zeros(a.shape, dtype=np.int64)
        act = ac.copy()
        self.fastclip(a, m, M, ac)
        self.clip(a, m, M, act)
        assert_array_strict_equal(ac, act)

    def test_simple_int64_inout(self):
        # Test native int32 input with double array min/max and int32 out.
        a = self._generate_int32_data(self.nr, self.nc)
        m = np.zeros(a.shape, np.float64)
        M = np.float64(1)
        ac = np.zeros(a.shape, dtype=np.int32)
        act = ac.copy()
        self.fastclip(a, m, M, out=ac, casting="unsafe")
        self.clip(a, m, M, act)
        assert_array_strict_equal(ac, act)

    def test_simple_int32_out(self):
        # Test native double input with scalar min/max and int out.
        a = self._generate_data(self.nr, self.nc)
        m = -1.0
        M = 2.0
        ac = np.zeros(a.shape, dtype=np.int32)
        act = ac.copy()
        self.fastclip(a, m, M, out=ac, casting="unsafe")
        self.clip(a, m, M, act)
        assert_array_strict_equal(ac, act)

    def test_simple_inplace_01(self):
        # Test native double input with array min/max in-place.
        a = self._generate_data(self.nr, self.nc)
        ac = a.copy()
        m = np.zeros(a.shape)
        M = 1.0
        self.fastclip(a, m, M, a)
        self.clip(a, m, M, ac)
        assert_array_strict_equal(a, ac)

    def test_simple_inplace_02(self):
        # Test native double input with scalar min/max in-place.
        a = self._generate_data(self.nr, self.nc)
        ac = a.copy()
        m = -0.5
        M = 0.6
        self.fastclip(a, m, M, a)
        self.clip(ac, m, M, ac)
        assert_array_strict_equal(a, ac)

    def test_noncontig_inplace(self):
        # Test non contiguous double input with double scalar min/max in-place.
        a = self._generate_data(self.nr * 2, self.nc * 3)
        a = a[::2, ::3]
        assert_(not a.flags['F_CONTIGUOUS'])
        assert_(not a.flags['C_CONTIGUOUS'])
        ac = a.copy()
        m = -0.5
        M = 0.6
        self.fastclip(a, m, M, a)
        self.clip(ac, m, M, ac)
        assert_array_equal(a, ac)

    def test_type_cast_01(self):
        # Test native double input with scalar min/max.
        a = self._generate_data(self.nr, self.nc)
        m = -0.5
        M = 0.6
        ac = self.fastclip(a, m, M)
        act = self.clip(a, m, M)
        assert_array_strict_equal(ac, act)

    def test_type_cast_02(self):
        # Test native int32 input with int32 scalar min/max.
        a = self._generate_int_data(self.nr, self.nc)
        a = a.astype(np.int32)
        m = -2
        M = 4
        ac = self.fastclip(a, m, M)
        act = self.clip(a, m, M)
        assert_array_strict_equal(ac, act)

    def test_type_cast_03(self):
        # Test native int32 input with float64 scalar min/max.
        a = self._generate_int32_data(self.nr, self.nc)
        m = -2
        M = 4
        ac = self.fastclip(a, np.float64(m), np.float64(M))
        act = self.clip(a, np.float64(m), np.float64(M))
        assert_array_strict_equal(ac, act)

    def test_type_cast_04(self):
        # Test native int32 input with float32 scalar min/max.
        a = self._generate_int32_data(self.nr, self.nc)
        m = np.float32(-2)
        M = np.float32(4)
        act = self.fastclip(a, m, M)
        ac = self.clip(a, m, M)
        assert_array_strict_equal(ac, act)

    def test_type_cast_05(self):
        # Test native int32 with double arrays min/max.
        a = self._generate_int_data(self.nr, self.nc)
        m = -0.5
        M = 1.
        ac = self.fastclip(a, m * np.zeros(a.shape), M)
        act = self.clip(a, m * np.zeros(a.shape), M)
        assert_array_strict_equal(ac, act)

    def test_type_cast_06(self):
        # Test native with NON native scalar min/max.
        a = self._generate_data(self.nr, self.nc)
        m = 0.5
        m_s = self._neg_byteorder(m)
        M = 1.
        act = self.clip(a, m_s, M)
        ac = self.fastclip(a, m_s, M)
        assert_array_strict_equal(ac, act)

    def test_type_cast_07(self):
        # Test NON native with native array min/max.
        a = self._generate_data(self.nr, self.nc)
        m = -0.5 * np.ones(a.shape)
        M = 1.
        a_s = self._neg_byteorder(a)
        assert_(not a_s.dtype.isnative)
        act = a_s.clip(m, M)
        ac = self.fastclip(a_s, m, M)
        assert_array_strict_equal(ac, act)

    def test_type_cast_08(self):
        # Test NON native with native scalar min/max.
        a = self._generate_data(self.nr, self.nc)
        m = -0.5
        M = 1.
        a_s = self._neg_byteorder(a)
        assert_(not a_s.dtype.isnative)
        ac = self.fastclip(a_s, m, M)
        act = a_s.clip(m, M)
        assert_array_strict_equal(ac, act)

    def test_type_cast_09(self):
        # Test native with NON native array min/max.
        a = self._generate_data(self.nr, self.nc)
        m = -0.5 * np.ones(a.shape)
        M = 1.
        m_s = self._neg_byteorder(m)
        assert_(not m_s.dtype.isnative)
        ac = self.fastclip(a, m_s, M)
        act = self.clip(a, m_s, M)
        assert_array_strict_equal(ac, act)

    def test_type_cast_10(self):
        # Test native int32 with float min/max and float out for output argument.
        a = self._generate_int_data(self.nr, self.nc)
        b = np.zeros(a.shape, dtype=np.float32)
        m = np.float32(-0.5)
        M = np.float32(1)
        act = self.clip(a, m, M, out=b)
        ac = self.fastclip(a, m, M, out=b)
        assert_array_strict_equal(ac, act)

    def test_type_cast_11(self):
        # Test non native with native scalar, min/max, out non native
        a = self._generate_non_native_data(self.nr, self.nc)
        b = a.copy()
        b = b.astype(b.dtype.newbyteorder('>'))
        bt = b.copy()
        m = -0.5
        M = 1.
        self.fastclip(a, m, M, out=b)
        self.clip(a, m, M, out=bt)
        assert_array_strict_equal(b, bt)

    def test_type_cast_12(self):
        # Test native int32 input and min/max and float out
        a = self._generate_int_data(self.nr, self.nc)
        b = np.zeros(a.shape, dtype=np.float32)
        m = np.int32(0)
        M = np.int32(1)
        act = self.clip(a, m, M, out=b)
        ac = self.fastclip(a, m, M, out=b)
        assert_array_strict_equal(ac, act)

    def test_clip_with_out_simple(self):
        # Test native double input with scalar min/max
        a = self._generate_data(self.nr, self.nc)
        m = -0.5
        M = 0.6
        ac = np.zeros(a.shape)
        act = np.zeros(a.shape)
        self.fastclip(a, m, M, ac)
        self.clip(a, m, M, act)
        assert_array_strict_equal(ac, act)

    def test_clip_with_out_simple2(self):
        # Test native int32 input with double min/max and int32 out
        a = self._generate_int32_data(self.nr, self.nc)
        m = np.float64(0)
        M = np.float64(2)
        ac = np.zeros(a.shape, dtype=np.int32)
        act = ac.copy()
        self.fastclip(a, m, M, out=ac, casting="unsafe")
        self.clip(a, m, M, act)
        assert_array_strict_equal(ac, act)

    def test_clip_with_out_simple_int32(self):
        # Test native int32 input with int32 scalar min/max and int64 out
        a = self._generate_int32_data(self.nr, self.nc)
        m = np.int32(-1)
        M = np.int32(1)
        ac = np.zeros(a.shape, dtype=np.int64)
        act = ac.copy()
        self.fastclip(a, m, M, ac)
        self.clip(a, m, M, act)
        assert_array_strict_equal(ac, act)

    def test_clip_with_out_array_int32(self):
        # Test native int32 input with double array min/max and int32 out
        a = self._generate_int32_data(self.nr, self.nc)
        m = np.zeros(a.shape, np.float64)
        M = np.float64(1)
        ac = np.zeros(a.shape, dtype=np.int32)
        act = ac.copy()
        self.fastclip(a, m, M, out=ac, casting="unsafe")
        self.clip(a, m, M, act)
        assert_array_strict_equal(ac, act)

    def test_clip_with_out_array_outint32(self):
        # Test native double input with scalar min/max and int out
        a = self._generate_data(self.nr, self.nc)
        m = -1.0
        M = 2.0
        ac = np.zeros(a.shape, dtype=np.int32)
        act = ac.copy()
        self.fastclip(a, m, M, out=ac, casting="unsafe")
        self.clip(a, m, M, act)
        assert_array_strict_equal(ac, act)

    def test_clip_with_out_transposed(self):
        # Test that the out argument works when transposed
        a = np.arange(16).reshape(4, 4)
        out = np.empty_like(a).T
        a.clip(4, 10, out=out)
        expected = self.clip(a, 4, 10)
        assert_array_equal(out, expected)

    def test_clip_with_out_memory_overlap(self):
        # Test that the out argument works when it has memory overlap
        a = np.arange(16).reshape(4, 4)
        ac = a.copy()
        a[:-1].clip(4, 10, out=a[1:])
        expected = self.clip(ac[:-1], 4, 10)
        assert_array_equal(a[1:], expected)

    def test_clip_inplace_array(self):
        # Test native double input with array min/max
        a = self._generate_data(self.nr, self.nc)
        ac = a.copy()
        m = np.zeros(a.shape)
        M = 1.0
        self.fastclip(a, m, M, a)
        self.clip(a, m, M, ac)
        assert_array_strict_equal(a, ac)

    def test_clip_inplace_simple(self):
        # Test native double input with scalar min/max
        a = self._generate_data(self.nr, self.nc)
        ac = a.copy()
        m = -0.5
        M = 0.6
        self.fastclip(a, m, M, a)
        self.clip(a, m, M, ac)
        assert_array_strict_equal(a, ac)

    def test_clip_func_takes_out(self):
        # Ensure that the clip() function takes an out=argument.
        a = self._generate_data(self.nr, self.nc)
        ac = a.copy()
        m = -0.5
        M = 0.6
        a2 = np.clip(a, m, M, out=a)
        self.clip(a, m, M, ac)
        assert_array_strict_equal(a2, ac)
        assert_(a2 is a)

    def test_clip_nan(self):
        d = np.arange(7.)
        assert_equal(d.clip(min=np.nan), np.nan)
        assert_equal(d.clip(max=np.nan), np.nan)
        assert_equal(d.clip(min=np.nan, max=np.nan), np.nan)
        assert_equal(d.clip(min=-2, max=np.nan), np.nan)
        assert_equal(d.clip(min=np.nan, max=10), np.nan)

    def test_object_clip(self):
        a = np.arange(10, dtype=object)
        actual = np.clip(a, 1, 5)
        expected = np.array([1, 1, 2, 3, 4, 5, 5, 5, 5, 5])
        assert actual.tolist() == expected.tolist()

    def test_clip_all_none(self):
        arr = np.arange(10, dtype=object)
        assert_equal(np.clip(arr, None, None), arr)
        assert_equal(np.clip(arr), arr)

    def test_clip_invalid_casting(self):
        a = np.arange(10, dtype=object)
        with assert_raises_regex(ValueError,
                                 'casting must be one of'):
            self.fastclip(a, 1, 8, casting="garbage")

    @pytest.mark.parametrize("amin, amax", [
        # two scalars
        (1, 0),
        # mix scalar and array
        (1, np.zeros(10)),
        # two arrays
        (np.ones(10), np.zeros(10)),
        ])
    def test_clip_value_min_max_flip(self, amin, amax):
        a = np.arange(10, dtype=np.int64)
        # requirement from ufunc_docstrings.py
        expected = np.minimum(np.maximum(a, amin), amax)
        actual = np.clip(a, amin, amax)
        assert_equal(actual, expected)

    @pytest.mark.parametrize("arr, amin, amax, exp", [
        # for a bug in npy_ObjectClip, based on a
        # case produced by hypothesis
        (np.zeros(10, dtype=object),
         0,
         -2**64 + 1,
         np.full(10, -2**64 + 1, dtype=object)),
        # for bugs in NPY_TIMEDELTA_MAX, based on a case
        # produced by hypothesis
        (np.zeros(10, dtype='m8') - 1,
         0,
         0,
         np.zeros(10, dtype='m8')),
    ])
    def test_clip_problem_cases(self, arr, amin, amax, exp):
        actual = np.clip(arr, amin, amax)
        assert_equal(actual, exp)

    @pytest.mark.parametrize("arr, amin, amax", [
        # problematic scalar nan case from hypothesis
        (np.zeros(10, dtype=np.int64),
         np.array(np.nan),
         np.zeros(10, dtype=np.int32)),
    ])
    def test_clip_scalar_nan_propagation(self, arr, amin, amax):
        # enforcement of scalar nan propagation for comparisons
        # called through clip()
        expected = np.minimum(np.maximum(arr, amin), amax)
        actual = np.clip(arr, amin, amax)
        assert_equal(actual, expected)

    @pytest.mark.xfail(reason="propagation doesn't match spec")
    @pytest.mark.parametrize("arr, amin, amax", [
        (np.array([1] * 10, dtype='m8'),
         np.timedelta64('NaT'),
         np.zeros(10, dtype=np.int32)),
    ])
    @pytest.mark.filterwarnings("ignore::DeprecationWarning")
    def test_NaT_propagation(self, arr, amin, amax):
        # NOTE: the expected function spec doesn't
        # propagate NaT, but clip() now does
        expected = np.minimum(np.maximum(arr, amin), amax)
        actual = np.clip(arr, amin, amax)
        assert_equal(actual, expected)

    @given(
        data=st.data(),
        arr=hynp.arrays(
            dtype=hynp.integer_dtypes() | hynp.floating_dtypes(),
            shape=hynp.array_shapes()
        )
    )
    def test_clip_property(self, data, arr):
        """A property-based test using Hypothesis.

        This aims for maximum generality: it could in principle generate *any*
        valid inputs to np.clip, and in practice generates much more varied
        inputs than human testers come up with.

        Because many of the inputs have tricky dependencies - compatible dtypes
        and mutually-broadcastable shapes - we use `st.data()` strategy draw
        values *inside* the test function, from strategies we construct based
        on previous values.  An alternative would be to define a custom strategy
        with `@st.composite`, but until we have duplicated code inline is fine.

        That accounts for most of the function; the actual test is just three
        lines to calculate and compare actual vs expected results!
        """
        numeric_dtypes = hynp.integer_dtypes() | hynp.floating_dtypes()
        # Generate shapes for the bounds which can be broadcast with each other
        # and with the base shape.  Below, we might decide to use scalar bounds,
        # but it's clearer to generate these shapes unconditionally in advance.
        in_shapes, result_shape = data.draw(
            hynp.mutually_broadcastable_shapes(
                num_shapes=2, base_shape=arr.shape
            )
        )
        # Scalar `nan` is deprecated due to the differing behaviour it shows.
        s = numeric_dtypes.flatmap(
            lambda x: hynp.from_dtype(x, allow_nan=False))
        amin = data.draw(s | hynp.arrays(dtype=numeric_dtypes,
            shape=in_shapes[0], elements={"allow_nan": False}))
        amax = data.draw(s | hynp.arrays(dtype=numeric_dtypes,
            shape=in_shapes[1], elements={"allow_nan": False}))

        # Then calculate our result and expected result and check that they're
        # equal!  See gh-12519 and gh-19457 for discussion deciding on this
        # property and the result_type argument.
        result = np.clip(arr, amin, amax)
        t = np.result_type(arr, amin, amax)
        expected = np.minimum(amax, np.maximum(arr, amin, dtype=t), dtype=t)
        assert result.dtype == t
        assert_array_equal(result, expected)

    def test_clip_min_max_args(self):
        arr = np.arange(5)

        assert_array_equal(np.clip(arr), arr)
        assert_array_equal(np.clip(arr, min=2, max=3), np.clip(arr, 2, 3))
        assert_array_equal(np.clip(arr, min=None, max=2),
                           np.clip(arr, None, 2))

        with assert_raises_regex(TypeError, "missing 1 required positional "
                                 "argument: 'a_max'"):
            np.clip(arr, 2)
        with assert_raises_regex(TypeError, "missing 1 required positional "
                                 "argument: 'a_min'"):
            np.clip(arr, a_max=2)
        msg = ("Passing `min` or `max` keyword argument when `a_min` and "
               "`a_max` are provided is forbidden.")
        with assert_raises_regex(ValueError, msg):
            np.clip(arr, 2, 3, max=3)
        with assert_raises_regex(ValueError, msg):
            np.clip(arr, 2, 3, min=2)

    @pytest.mark.parametrize("dtype,min,max", [
        ("int32", -2**32 - 1, 2**32),
        ("int32", -2**320, None),
        ("int32", None, 2**300),
        ("int32", -1000, 2**32),
        ("int32", -2**32 - 1, 1000),
        ("uint8", -1, 129),
    ])
    def test_out_of_bound_pyints(self, dtype, min, max):
        a = np.arange(10000).astype(dtype)
        # Check min only
        c = np.clip(a, min=min, max=max)
        assert not np.may_share_memory(a, c)
        assert c.dtype == a.dtype
        if min is not None:
            assert (c >= min).all()
        if max is not None:
            assert (c <= max).all()

class TestAllclose:
    rtol = 1e-5
    atol = 1e-8

    def setup_method(self):
        self.olderr = np.seterr(invalid='ignore')

    def teardown_method(self):
        np.seterr(**self.olderr)

    def tst_allclose(self, x, y):
        assert_(np.allclose(x, y), f"{x} and {y} not close")

    def tst_not_allclose(self, x, y):
        assert_(not np.allclose(x, y), f"{x} and {y} shouldn't be close")

    def test_ip_allclose(self):
        # Parametric test factory.
        arr = np.array([100, 1000])
        aran = np.arange(125).reshape((5, 5, 5))

        atol = self.atol
        rtol = self.rtol

        data = [([1, 0], [1, 0]),
                ([atol], [0]),
                ([1], [1 + rtol + atol]),
                (arr, arr + arr * rtol),
                (arr, arr + arr * rtol + atol * 2),
                (aran, aran + aran * rtol),
                (np.inf, np.inf),
                (np.inf, [np.inf])]

        for (x, y) in data:
            self.tst_allclose(x, y)

    def test_ip_not_allclose(self):
        # Parametric test factory.
        aran = np.arange(125).reshape((5, 5, 5))

        atol = self.atol
        rtol = self.rtol

        data = [([np.inf, 0], [1, np.inf]),
                ([np.inf, 0], [1, 0]),
                ([np.inf, np.inf], [1, np.inf]),
                ([np.inf, np.inf], [1, 0]),
                ([-np.inf, 0], [np.inf, 0]),
                ([np.nan, 0], [np.nan, 0]),
                ([atol * 2], [0]),
                ([1], [1 + rtol + atol * 2]),
                (aran, aran + aran * atol + atol * 2),
                (np.array([np.inf, 1]), np.array([0, np.inf]))]

        for (x, y) in data:
            self.tst_not_allclose(x, y)

    def test_no_parameter_modification(self):
        x = np.array([np.inf, 1])
        y = np.array([0, np.inf])
        np.allclose(x, y)
        assert_array_equal(x, np.array([np.inf, 1]))
        assert_array_equal(y, np.array([0, np.inf]))

    def test_min_int(self):
        # Could make problems because of abs(min_int) == min_int
        min_int = np.iinfo(np.int_).min
        a = np.array([min_int], dtype=np.int_)
        assert_(np.allclose(a, a))

    def test_equalnan(self):
        x = np.array([1.0, np.nan])
        assert_(np.allclose(x, x, equal_nan=True))

    def test_return_class_is_ndarray(self):
        # Issue gh-6475
        # Check that allclose does not preserve subtypes
        class Foo(np.ndarray):
            def __new__(cls, *args, **kwargs):
                return np.array(*args, **kwargs).view(cls)

        a = Foo([1])
        assert_(type(np.allclose(a, a)) is bool)


class TestIsclose:
    rtol = 1e-5
    atol = 1e-8

    def _setup(self):
        atol = self.atol
        rtol = self.rtol
        arr = np.array([100, 1000])
        aran = np.arange(125).reshape((5, 5, 5))

        self.all_close_tests = [
                ([1, 0], [1, 0]),
                ([atol], [0]),
                ([1], [1 + rtol + atol]),
                (arr, arr + arr * rtol),
                (arr, arr + arr * rtol + atol),
                (aran, aran + aran * rtol),
                (np.inf, np.inf),
                (np.inf, [np.inf]),
                ([np.inf, -np.inf], [np.inf, -np.inf]),
                ]
        self.none_close_tests = [
                ([np.inf, 0], [1, np.inf]),
                ([np.inf, -np.inf], [1, 0]),
                ([np.inf, np.inf], [1, -np.inf]),
                ([np.inf, np.inf], [1, 0]),
                ([np.nan, 0], [np.nan, -np.inf]),
                ([atol * 2], [0]),
                ([1], [1 + rtol + atol * 2]),
                (aran, aran + rtol * 1.1 * aran + atol * 1.1),
                (np.array([np.inf, 1]), np.array([0, np.inf])),
                ]
        self.some_close_tests = [
                ([np.inf, 0], [np.inf, atol * 2]),
                ([atol, 1, 1e6 * (1 + 2 * rtol) + atol], [0, np.nan, 1e6]),
                (np.arange(3), [0, 1, 2.1]),
                (np.nan, [np.nan, np.nan, np.nan]),
                ([0], [atol, np.inf, -np.inf, np.nan]),
                (0, [atol, np.inf, -np.inf, np.nan]),
                ]
        self.some_close_results = [
                [True, False],
                [True, False, False],
                [True, True, False],
                [False, False, False],
                [True, False, False, False],
                [True, False, False, False],
                ]

    def test_ip_isclose(self):
        self._setup()
        tests = self.some_close_tests
        results = self.some_close_results
        for (x, y), result in zip(tests, results):
            assert_array_equal(np.isclose(x, y), result)

        x = np.array([2.1, 2.1, 2.1, 2.1, 5, np.nan])
        y = np.array([2, 2, 2, 2, np.nan, 5])
        atol = [0.11, 0.09, 1e-8, 1e-8, 1, 1]
        rtol = [1e-8, 1e-8, 0.06, 0.04, 1, 1]
        expected = np.array([True, False, True, False, False, False])
        assert_array_equal(np.isclose(x, y, rtol=rtol, atol=atol), expected)

        message = "operands could not be broadcast together..."
        atol = np.array([1e-8, 1e-8])
        with assert_raises(ValueError, msg=message):
            np.isclose(x, y, atol=atol)

        rtol = np.array([1e-5, 1e-5])
        with assert_raises(ValueError, msg=message):
            np.isclose(x, y, rtol=rtol)

    def test_nep50_isclose(self):
        below_one = float(1. - np.finfo('f8').eps)
        f32 = np.array(below_one, 'f4')  # This is just 1 at float32 precision
        assert f32 > np.array(below_one)
        # NEP 50 broadcasting of python scalars
        assert f32 == below_one
        # Test that it works for isclose arguments too (and that those fail if
        # one uses a numpy float64).
        assert np.isclose(f32, below_one, atol=0, rtol=0)
        assert np.isclose(f32, np.float32(0), atol=below_one)
        assert np.isclose(f32, 2, atol=0, rtol=below_one / 2)
        assert not np.isclose(f32, np.float64(below_one), atol=0, rtol=0)
        assert not np.isclose(f32, np.float32(0), atol=np.float64(below_one))
        assert not np.isclose(f32, 2, atol=0, rtol=np.float64(below_one / 2))

    def tst_all_isclose(self, x, y):
        assert_(np.all(np.isclose(x, y)), f"{x} and {y} not close")

    def tst_none_isclose(self, x, y):
        msg = "%s and %s shouldn't be close"
        assert_(not np.any(np.isclose(x, y)), msg % (x, y))

    def tst_isclose_allclose(self, x, y):
        msg = "isclose.all() and allclose aren't same for %s and %s"
        msg2 = "isclose and allclose aren't same for %s and %s"
        if np.isscalar(x) and np.isscalar(y):
            assert_(np.isclose(x, y) == np.allclose(x, y), msg=msg2 % (x, y))
        else:
            assert_array_equal(np.isclose(x, y).all(), np.allclose(x, y), msg % (x, y))

    def test_ip_all_isclose(self):
        self._setup()
        for (x, y) in self.all_close_tests:
            self.tst_all_isclose(x, y)

        x = np.array([2.3, 3.6, 4.4, np.nan])
        y = np.array([2, 3, 4, np.nan])
        atol = [0.31, 0, 0, 1]
        rtol = [0, 0.21, 0.11, 1]
        assert np.allclose(x, y, atol=atol, rtol=rtol, equal_nan=True)
        assert not np.allclose(x, y, atol=0.1, rtol=0.1, equal_nan=True)

        # Show that gh-14330 is resolved
        assert np.allclose([1, 2, float('nan')], [1, 2, float('nan')],
                           atol=[1, 1, 1], equal_nan=True)

    def test_ip_none_isclose(self):
        self._setup()
        for (x, y) in self.none_close_tests:
            self.tst_none_isclose(x, y)

    def test_ip_isclose_allclose(self):
        self._setup()
        tests = (self.all_close_tests + self.none_close_tests +
                 self.some_close_tests)
        for (x, y) in tests:
            self.tst_isclose_allclose(x, y)

    def test_equal_nan(self):
        assert_array_equal(np.isclose(np.nan, np.nan, equal_nan=True), [True])
        arr = np.array([1.0, np.nan])
        assert_array_equal(np.isclose(arr, arr, equal_nan=True), [True, True])

    def test_masked_arrays(self):
        # Make sure to test the output type when arguments are interchanged.

        x = np.ma.masked_where([True, True, False], np.arange(3))
        assert_(type(x) is type(np.isclose(2, x)))
        assert_(type(x) is type(np.isclose(x, 2)))

        x = np.ma.masked_where([True, True, False], [np.nan, np.inf, np.nan])
        assert_(type(x) is type(np.isclose(np.inf, x)))
        assert_(type(x) is type(np.isclose(x, np.inf)))

        x = np.ma.masked_where([True, True, False], [np.nan, np.nan, np.nan])
        y = np.isclose(np.nan, x, equal_nan=True)
        assert_(type(x) is type(y))
        # Ensure that the mask isn't modified...
        assert_array_equal([True, True, False], y.mask)
        y = np.isclose(x, np.nan, equal_nan=True)
        assert_(type(x) is type(y))
        # Ensure that the mask isn't modified...
        assert_array_equal([True, True, False], y.mask)

        x = np.ma.masked_where([True, True, False], [np.nan, np.nan, np.nan])
        y = np.isclose(x, x, equal_nan=True)
        assert_(type(x) is type(y))
        # Ensure that the mask isn't modified...
        assert_array_equal([True, True, False], y.mask)

    def test_scalar_return(self):
        assert_(np.isscalar(np.isclose(1, 1)))

    def test_no_parameter_modification(self):
        x = np.array([np.inf, 1])
        y = np.array([0, np.inf])
        np.isclose(x, y)
        assert_array_equal(x, np.array([np.inf, 1]))
        assert_array_equal(y, np.array([0, np.inf]))

    def test_non_finite_scalar(self):
        # GH7014, when two scalars are compared the output should also be a
        # scalar
        assert_(np.isclose(np.inf, -np.inf) is np.False_)
        assert_(np.isclose(0, np.inf) is np.False_)
        assert_(type(np.isclose(0, np.inf)) is np.bool)

    def test_timedelta(self):
        # Allclose currently works for timedelta64 as long as `atol` is
        # an integer or also a timedelta64
        a = np.array([[1, 2, 3, "NaT"]], dtype="m8[ns]")
        assert np.isclose(a, a, atol=0, equal_nan=True).all()
        assert np.isclose(a, a, atol=np.timedelta64(1, "ns"), equal_nan=True).all()
        assert np.allclose(a, a, atol=0, equal_nan=True)
        assert np.allclose(a, a, atol=np.timedelta64(1, "ns"), equal_nan=True)

    def test_tol_warnings(self):
        a = np.array([1, 2, 3])
        b = np.array([np.inf, np.nan, 1])

        for i in b:
            for j in b:
                # Making sure that i and j are not both numbers, because that won't create a warning
                if (i == 1) and (j == 1):
                    continue

                with warnings.catch_warnings(record=True) as w:

                    warnings.simplefilter("always")
                    c = np.isclose(a, a, atol=i, rtol=j)
                    assert len(w) == 1
                    assert issubclass(w[-1].category, RuntimeWarning)
                    assert f"One of rtol or atol is not valid, atol: {i}, rtol: {j}" in str(w[-1].message)


class TestStdVar:
    def setup_method(self):
        self.A = np.array([1, -1, 1, -1])
        self.real_var = 1

    def test_basic(self):
        assert_almost_equal(np.var(self.A), self.real_var)
        assert_almost_equal(np.std(self.A)**2, self.real_var)

    def test_scalars(self):
        assert_equal(np.var(1), 0)
        assert_equal(np.std(1), 0)

    def test_ddof1(self):
        assert_almost_equal(np.var(self.A, ddof=1),
                            self.real_var * len(self.A) / (len(self.A) - 1))
        assert_almost_equal(np.std(self.A, ddof=1)**2,
                            self.real_var * len(self.A) / (len(self.A) - 1))

    def test_ddof2(self):
        assert_almost_equal(np.var(self.A, ddof=2),
                            self.real_var * len(self.A) / (len(self.A) - 2))
        assert_almost_equal(np.std(self.A, ddof=2)**2,
                            self.real_var * len(self.A) / (len(self.A) - 2))

    def test_correction(self):
        assert_almost_equal(
            np.var(self.A, correction=1), np.var(self.A, ddof=1)
        )
        assert_almost_equal(
            np.std(self.A, correction=1), np.std(self.A, ddof=1)
        )

        err_msg = "ddof and correction can't be provided simultaneously."

        with assert_raises_regex(ValueError, err_msg):
            np.var(self.A, ddof=1, correction=0)

        with assert_raises_regex(ValueError, err_msg):
            np.std(self.A, ddof=1, correction=1)

    def test_out_scalar(self):
        d = np.arange(10)
        out = np.array(0.)
        r = np.std(d, out=out)
        assert_(r is out)
        assert_array_equal(r, out)
        r = np.var(d, out=out)
        assert_(r is out)
        assert_array_equal(r, out)
        r = np.mean(d, out=out)
        assert_(r is out)
        assert_array_equal(r, out)


class TestStdVarComplex:
    def test_basic(self):
        A = np.array([1, 1.j, -1, -1.j])
        real_var = 1
        assert_almost_equal(np.var(A), real_var)
        assert_almost_equal(np.std(A)**2, real_var)

    def test_scalars(self):
        assert_equal(np.var(1j), 0)
        assert_equal(np.std(1j), 0)


class TestCreationFuncs:
    # Test ones, zeros, empty and full.

    def setup_method(self):
        dtypes = {np.dtype(tp) for tp in itertools.chain(*sctypes.values())}
        # void, bytes, str
        variable_sized = {tp for tp in dtypes if tp.str.endswith('0')}
        keyfunc = lambda dtype: dtype.str
        self.dtypes = sorted(dtypes - variable_sized |
                             {np.dtype(tp.str.replace("0", str(i)))
                              for tp in variable_sized for i in range(1, 10)},
                             key=keyfunc)
        self.dtypes += [type(dt) for dt in sorted(dtypes, key=keyfunc)]
        self.orders = {'C': 'c_contiguous', 'F': 'f_contiguous'}
        self.ndims = 10

    def check_function(self, func, fill_value=None):
        par = ((0, 1, 2),
               range(self.ndims),
               self.orders,
               self.dtypes)
        fill_kwarg = {}
        if fill_value is not None:
            fill_kwarg = {'fill_value': fill_value}

        for size, ndims, order, dtype in itertools.product(*par):
            shape = ndims * [size]

            is_void = dtype is np.dtypes.VoidDType or (
                isinstance(dtype, np.dtype) and dtype.str.startswith('|V'))

            # do not fill void type
            if fill_kwarg and is_void:
                continue

            arr = func(shape, order=order, dtype=dtype,
                       **fill_kwarg)

            if isinstance(dtype, np.dtype):
                assert_equal(arr.dtype, dtype)
            elif isinstance(dtype, type(np.dtype)):
                if dtype in (np.dtypes.StrDType, np.dtypes.BytesDType):
                    dtype_str = np.dtype(dtype.type).str.replace('0', '1')
                    assert_equal(arr.dtype, np.dtype(dtype_str))
                else:
                    assert_equal(arr.dtype, np.dtype(dtype.type))
            assert_(getattr(arr.flags, self.orders[order]))

            if fill_value is not None:
                if arr.dtype.str.startswith('|S'):
                    val = str(fill_value)
                else:
                    val = fill_value
                assert_equal(arr, dtype.type(val))

    def test_zeros(self):
        self.check_function(np.zeros)

    def test_ones(self):
        self.check_function(np.ones)

    def test_empty(self):
        self.check_function(np.empty)

    def test_full(self):
        self.check_function(np.full, 0)
        self.check_function(np.full, 1)

    @pytest.mark.skipif(not HAS_REFCOUNT, reason="Python lacks refcounts")
    def test_for_reference_leak(self):
        # Make sure we have an object for reference
        dim = 1
        beg = sys.getrefcount(dim)
        np.zeros([dim] * 10)
        assert_(sys.getrefcount(dim) == beg)
        np.ones([dim] * 10)
        assert_(sys.getrefcount(dim) == beg)
        np.empty([dim] * 10)
        assert_(sys.getrefcount(dim) == beg)
        np.full([dim] * 10, 0)
        assert_(sys.getrefcount(dim) == beg)


class TestLikeFuncs:
    '''Test ones_like, zeros_like, empty_like and full_like'''

    def setup_method(self):
        self.data = [
                # Array scalars
                (np.array(3.), None),
                (np.array(3), 'f8'),
                # 1D arrays
                (np.arange(6, dtype='f4'), None),
                (np.arange(6), 'c16'),
                # 2D C-layout arrays
                (np.arange(6).reshape(2, 3), None),
                (np.arange(6).reshape(3, 2), 'i1'),
                # 2D F-layout arrays
                (np.arange(6).reshape((2, 3), order='F'), None),
                (np.arange(6).reshape((3, 2), order='F'), 'i1'),
                # 3D C-layout arrays
                (np.arange(24).reshape(2, 3, 4), None),
                (np.arange(24).reshape(4, 3, 2), 'f4'),
                # 3D F-layout arrays
                (np.arange(24).reshape((2, 3, 4), order='F'), None),
                (np.arange(24).reshape((4, 3, 2), order='F'), 'f4'),
                # 3D non-C/F-layout arrays
                (np.arange(24).reshape(2, 3, 4).swapaxes(0, 1), None),
                (np.arange(24).reshape(4, 3, 2).swapaxes(0, 1), '?'),
                     ]
        self.shapes = [(), (5,), (5, 6,), (5, 6, 7,)]

    def compare_array_value(self, dz, value, fill_value):
        if value is not None:
            if fill_value:
                # Conversion is close to what np.full_like uses
                # but we  may want to convert directly in the future
                # which may result in errors (where this does not).
                z = np.array(value).astype(dz.dtype)
                assert_(np.all(dz == z))
            else:
                assert_(np.all(dz == value))

    def check_like_function(self, like_function, value, fill_value=False):
        if fill_value:
            fill_kwarg = {'fill_value': value}
        else:
            fill_kwarg = {}
        for d, dtype in self.data:
            # default (K) order, dtype
            dz = like_function(d, dtype=dtype, **fill_kwarg)
            assert_equal(dz.shape, d.shape)
            assert_equal(np.array(dz.strides) * d.dtype.itemsize,
                         np.array(d.strides) * dz.dtype.itemsize)
            assert_equal(d.flags.c_contiguous, dz.flags.c_contiguous)
            assert_equal(d.flags.f_contiguous, dz.flags.f_contiguous)
            if dtype is None:
                assert_equal(dz.dtype, d.dtype)
            else:
                assert_equal(dz.dtype, np.dtype(dtype))
            self.compare_array_value(dz, value, fill_value)

            # C order, default dtype
            dz = like_function(d, order='C', dtype=dtype, **fill_kwarg)
            assert_equal(dz.shape, d.shape)
            assert_(dz.flags.c_contiguous)
            if dtype is None:
                assert_equal(dz.dtype, d.dtype)
            else:
                assert_equal(dz.dtype, np.dtype(dtype))
            self.compare_array_value(dz, value, fill_value)

            # F order, default dtype
            dz = like_function(d, order='F', dtype=dtype, **fill_kwarg)
            assert_equal(dz.shape, d.shape)
            assert_(dz.flags.f_contiguous)
            if dtype is None:
                assert_equal(dz.dtype, d.dtype)
            else:
                assert_equal(dz.dtype, np.dtype(dtype))
            self.compare_array_value(dz, value, fill_value)

            # A order
            dz = like_function(d, order='A', dtype=dtype, **fill_kwarg)
            assert_equal(dz.shape, d.shape)
            if d.flags.f_contiguous:
                assert_(dz.flags.f_contiguous)
            else:
                assert_(dz.flags.c_contiguous)
            if dtype is None:
                assert_equal(dz.dtype, d.dtype)
            else:
                assert_equal(dz.dtype, np.dtype(dtype))
            self.compare_array_value(dz, value, fill_value)

            # Test the 'shape' parameter
            for s in self.shapes:
                for o in 'CFA':
                    sz = like_function(d, dtype=dtype, shape=s, order=o,
                                       **fill_kwarg)
                    assert_equal(sz.shape, s)
                    if dtype is None:
                        assert_equal(sz.dtype, d.dtype)
                    else:
                        assert_equal(sz.dtype, np.dtype(dtype))
                    if o == 'C' or (o == 'A' and d.flags.c_contiguous):
                        assert_(sz.flags.c_contiguous)
                    elif o == 'F' or (o == 'A' and d.flags.f_contiguous):
                        assert_(sz.flags.f_contiguous)
                    self.compare_array_value(sz, value, fill_value)

                if (d.ndim != len(s)):
                    assert_equal(np.argsort(like_function(d, dtype=dtype,
                                                          shape=s, order='K',
                                                          **fill_kwarg).strides),
                                 np.argsort(np.empty(s, dtype=dtype,
                                                     order='C').strides))
                else:
                    assert_equal(np.argsort(like_function(d, dtype=dtype,
                                                          shape=s, order='K',
                                                          **fill_kwarg).strides),
                                 np.argsort(d.strides))

        # Test the 'subok' parameter
        class MyNDArray(np.ndarray):
            pass

        a = np.array([[1, 2], [3, 4]]).view(MyNDArray)

        b = like_function(a, **fill_kwarg)
        assert_(type(b) is MyNDArray)

        b = like_function(a, subok=False, **fill_kwarg)
        assert_(type(b) is not MyNDArray)

        # Test invalid dtype
        with assert_raises(TypeError):
            a = np.array(b"abc")
            like_function(a, dtype="S-1", **fill_kwarg)

    def test_ones_like(self):
        self.check_like_function(np.ones_like, 1)

    def test_zeros_like(self):
        self.check_like_function(np.zeros_like, 0)

    def test_empty_like(self):
        self.check_like_function(np.empty_like, None)

    def test_filled_like(self):
        self.check_like_function(np.full_like, 0, True)
        self.check_like_function(np.full_like, 1, True)
        # Large integers may overflow, but using int64 is OK (casts)
        # see also gh-27075
        with pytest.raises(OverflowError):
            np.full_like(np.ones(3, dtype=np.int8), 1000)
        self.check_like_function(np.full_like, np.int64(1000), True)
        self.check_like_function(np.full_like, 123.456, True)
        # Inf to integer casts cause invalid-value errors: ignore them.
        with np.errstate(invalid="ignore"):
            self.check_like_function(np.full_like, np.inf, True)

    @pytest.mark.parametrize('likefunc', [np.empty_like, np.full_like,
                                          np.zeros_like, np.ones_like])
    @pytest.mark.parametrize('dtype', [str, bytes])
    def test_dtype_str_bytes(self, likefunc, dtype):
        # Regression test for gh-19860
        a = np.arange(16).reshape(2, 8)
        b = a[:, ::2]  # Ensure b is not contiguous.
        kwargs = {'fill_value': ''} if likefunc == np.full_like else {}
        result = likefunc(b, dtype=dtype, **kwargs)
        if dtype == str:
            assert result.strides == (16, 4)
        else:
            # dtype is bytes
            assert result.strides == (4, 1)


class TestCorrelate:
    def _setup(self, dt):
        self.x = np.array([1, 2, 3, 4, 5], dtype=dt)
        self.xs = np.arange(1, 20)[::3]
        self.y = np.array([-1, -2, -3], dtype=dt)
        self.z1 = np.array([-3., -8., -14., -20., -26., -14., -5.], dtype=dt)
        self.z1_4 = np.array([-2., -5., -8., -11., -14., -5.], dtype=dt)
        self.z1r = np.array([-15., -22., -22., -16., -10., -4., -1.], dtype=dt)
        self.z2 = np.array([-5., -14., -26., -20., -14., -8., -3.], dtype=dt)
        self.z2r = np.array([-1., -4., -10., -16., -22., -22., -15.], dtype=dt)
        self.zs = np.array([-3., -14., -30., -48., -66., -84.,
                           -102., -54., -19.], dtype=dt)

    def test_float(self):
        self._setup(float)
        z = np.correlate(self.x, self.y, 'full')
        assert_array_almost_equal(z, self.z1)
        z = np.correlate(self.x, self.y[:-1], 'full')
        assert_array_almost_equal(z, self.z1_4)
        z = np.correlate(self.y, self.x, 'full')
        assert_array_almost_equal(z, self.z2)
        z = np.correlate(self.x[::-1], self.y, 'full')
        assert_array_almost_equal(z, self.z1r)
        z = np.correlate(self.y, self.x[::-1], 'full')
        assert_array_almost_equal(z, self.z2r)
        z = np.correlate(self.xs, self.y, 'full')
        assert_array_almost_equal(z, self.zs)

    def test_object(self):
        self._setup(Decimal)
        z = np.correlate(self.x, self.y, 'full')
        assert_array_almost_equal(z, self.z1)
        z = np.correlate(self.y, self.x, 'full')
        assert_array_almost_equal(z, self.z2)

    def test_no_overwrite(self):
        d = np.ones(100)
        k = np.ones(3)
        np.correlate(d, k)
        assert_array_equal(d, np.ones(100))
        assert_array_equal(k, np.ones(3))

    def test_complex(self):
        x = np.array([1, 2, 3, 4 + 1j], dtype=complex)
        y = np.array([-1, -2j, 3 + 1j], dtype=complex)
        r_z = np.array([3 - 1j, 6, 8 + 1j, 11 + 5j, -5 + 8j, -4 - 1j], dtype=complex)
        r_z = r_z[::-1].conjugate()
        z = np.correlate(y, x, mode='full')
        assert_array_almost_equal(z, r_z)

    def test_zero_size(self):
        with pytest.raises(ValueError):
            np.correlate(np.array([]), np.ones(1000), mode='full')
        with pytest.raises(ValueError):
            np.correlate(np.ones(1000), np.array([]), mode='full')

    def test_mode(self):
        d = np.ones(100)
        k = np.ones(3)
        default_mode = np.correlate(d, k, mode='valid')
        with assert_raises(ValueError):
            np.correlate(d, k, mode='v')
        # integer mode
        with assert_raises(ValueError):
            np.correlate(d, k, mode=-1)
        # assert_array_equal(np.correlate(d, k, mode=), default_mode)
        # illegal arguments
        with assert_raises(TypeError):
            np.correlate(d, k, mode=None)


class TestConvolve:
    def test_object(self):
        d = [1.] * 100
        k = [1.] * 3
        assert_array_almost_equal(np.convolve(d, k)[2:-2], np.full(98, 3))

    def test_no_overwrite(self):
        d = np.ones(100)
        k = np.ones(3)
        np.convolve(d, k)
        assert_array_equal(d, np.ones(100))
        assert_array_equal(k, np.ones(3))

    def test_mode(self):
        d = np.ones(100)
        k = np.ones(3)
        default_mode = np.convolve(d, k, mode='full')
        with assert_raises(ValueError):
            np.convolve(d, k, mode='f')
        # integer mode
        with assert_raises(ValueError):
            np.convolve(d, k, mode=-1)
        assert_array_equal(np.convolve(d, k, mode=2), default_mode)
        # illegal arguments
        with assert_raises(TypeError):
            np.convolve(d, k, mode=None)


class TestArgwhere:

    @pytest.mark.parametrize('nd', [0, 1, 2])
    def test_nd(self, nd):
        # get an nd array with multiple elements in every dimension
        x = np.empty((2,) * nd, bool)

        # none
        x[...] = False
        assert_equal(np.argwhere(x).shape, (0, nd))

        # only one
        x[...] = False
        x.flat[0] = True
        assert_equal(np.argwhere(x).shape, (1, nd))

        # all but one
        x[...] = True
        x.flat[0] = False
        assert_equal(np.argwhere(x).shape, (x.size - 1, nd))

        # all
        x[...] = True
        assert_equal(np.argwhere(x).shape, (x.size, nd))

    def test_2D(self):
        x = np.arange(6).reshape((2, 3))
        assert_array_equal(np.argwhere(x > 1),
                           [[0, 2],
                            [1, 0],
                            [1, 1],
                            [1, 2]])

    def test_list(self):
        assert_equal(np.argwhere([4, 0, 2, 1, 3]), [[0], [2], [3], [4]])


class TestRoll:
    def test_roll1d(self):
        x = np.arange(10)
        xr = np.roll(x, 2)
        assert_equal(xr, np.array([8, 9, 0, 1, 2, 3, 4, 5, 6, 7]))

    def test_roll2d(self):
        x2 = np.reshape(np.arange(10), (2, 5))
        x2r = np.roll(x2, 1)
        assert_equal(x2r, np.array([[9, 0, 1, 2, 3], [4, 5, 6, 7, 8]]))

        x2r = np.roll(x2, 1, axis=0)
        assert_equal(x2r, np.array([[5, 6, 7, 8, 9], [0, 1, 2, 3, 4]]))

        x2r = np.roll(x2, 1, axis=1)
        assert_equal(x2r, np.array([[4, 0, 1, 2, 3], [9, 5, 6, 7, 8]]))

        # Roll multiple axes at once.
        x2r = np.roll(x2, 1, axis=(0, 1))
        assert_equal(x2r, np.array([[9, 5, 6, 7, 8], [4, 0, 1, 2, 3]]))

        x2r = np.roll(x2, (1, 0), axis=(0, 1))
        assert_equal(x2r, np.array([[5, 6, 7, 8, 9], [0, 1, 2, 3, 4]]))

        x2r = np.roll(x2, (-1, 0), axis=(0, 1))
        assert_equal(x2r, np.array([[5, 6, 7, 8, 9], [0, 1, 2, 3, 4]]))

        x2r = np.roll(x2, (0, 1), axis=(0, 1))
        assert_equal(x2r, np.array([[4, 0, 1, 2, 3], [9, 5, 6, 7, 8]]))

        x2r = np.roll(x2, (0, -1), axis=(0, 1))
        assert_equal(x2r, np.array([[1, 2, 3, 4, 0], [6, 7, 8, 9, 5]]))

        x2r = np.roll(x2, (1, 1), axis=(0, 1))
        assert_equal(x2r, np.array([[9, 5, 6, 7, 8], [4, 0, 1, 2, 3]]))

        x2r = np.roll(x2, (-1, -1), axis=(0, 1))
        assert_equal(x2r, np.array([[6, 7, 8, 9, 5], [1, 2, 3, 4, 0]]))

        # Roll the same axis multiple times.
        x2r = np.roll(x2, 1, axis=(0, 0))
        assert_equal(x2r, np.array([[0, 1, 2, 3, 4], [5, 6, 7, 8, 9]]))

        x2r = np.roll(x2, 1, axis=(1, 1))
        assert_equal(x2r, np.array([[3, 4, 0, 1, 2], [8, 9, 5, 6, 7]]))

        # Roll more than one turn in either direction.
        x2r = np.roll(x2, 6, axis=1)
        assert_equal(x2r, np.array([[4, 0, 1, 2, 3], [9, 5, 6, 7, 8]]))

        x2r = np.roll(x2, -4, axis=1)
        assert_equal(x2r, np.array([[4, 0, 1, 2, 3], [9, 5, 6, 7, 8]]))

    def test_roll_empty(self):
        x = np.array([])
        assert_equal(np.roll(x, 1), np.array([]))

    def test_roll_unsigned_shift(self):
        x = np.arange(4)
        shift = np.uint16(2)
        assert_equal(np.roll(x, shift), np.roll(x, 2))

        shift = np.uint64(2**63 + 2)
        assert_equal(np.roll(x, shift), np.roll(x, 2))

    def test_roll_big_int(self):
        x = np.arange(4)
        assert_equal(np.roll(x, 2**100), x)


class TestRollaxis:

    # expected shape indexed by (axis, start) for array of
    # shape (1, 2, 3, 4)
    tgtshape = {(0, 0): (1, 2, 3, 4), (0, 1): (1, 2, 3, 4),
                (0, 2): (2, 1, 3, 4), (0, 3): (2, 3, 1, 4),
                (0, 4): (2, 3, 4, 1),
                (1, 0): (2, 1, 3, 4), (1, 1): (1, 2, 3, 4),
                (1, 2): (1, 2, 3, 4), (1, 3): (1, 3, 2, 4),
                (1, 4): (1, 3, 4, 2),
                (2, 0): (3, 1, 2, 4), (2, 1): (1, 3, 2, 4),
                (2, 2): (1, 2, 3, 4), (2, 3): (1, 2, 3, 4),
                (2, 4): (1, 2, 4, 3),
                (3, 0): (4, 1, 2, 3), (3, 1): (1, 4, 2, 3),
                (3, 2): (1, 2, 4, 3), (3, 3): (1, 2, 3, 4),
                (3, 4): (1, 2, 3, 4)}

    def test_exceptions(self):
        a = np.arange(1 * 2 * 3 * 4).reshape(1, 2, 3, 4)
        assert_raises(AxisError, np.rollaxis, a, -5, 0)
        assert_raises(AxisError, np.rollaxis, a, 0, -5)
        assert_raises(AxisError, np.rollaxis, a, 4, 0)
        assert_raises(AxisError, np.rollaxis, a, 0, 5)

    def test_results(self):
        a = np.arange(1 * 2 * 3 * 4).reshape(1, 2, 3, 4).copy()
        aind = np.indices(a.shape)
        assert_(a.flags['OWNDATA'])
        for (i, j) in self.tgtshape:
            # positive axis, positive start
            res = np.rollaxis(a, axis=i, start=j)
            i0, i1, i2, i3 = aind[np.array(res.shape) - 1]
            assert_(np.all(res[i0, i1, i2, i3] == a))
            assert_(res.shape == self.tgtshape[(i, j)], str((i, j)))
            assert_(not res.flags['OWNDATA'])

            # negative axis, positive start
            ip = i + 1
            res = np.rollaxis(a, axis=-ip, start=j)
            i0, i1, i2, i3 = aind[np.array(res.shape) - 1]
            assert_(np.all(res[i0, i1, i2, i3] == a))
            assert_(res.shape == self.tgtshape[(4 - ip, j)])
            assert_(not res.flags['OWNDATA'])

            # positive axis, negative start
            jp = j + 1 if j < 4 else j
            res = np.rollaxis(a, axis=i, start=-jp)
            i0, i1, i2, i3 = aind[np.array(res.shape) - 1]
            assert_(np.all(res[i0, i1, i2, i3] == a))
            assert_(res.shape == self.tgtshape[(i, 4 - jp)])
            assert_(not res.flags['OWNDATA'])

            # negative axis, negative start
            ip = i + 1
            jp = j + 1 if j < 4 else j
            res = np.rollaxis(a, axis=-ip, start=-jp)
            i0, i1, i2, i3 = aind[np.array(res.shape) - 1]
            assert_(np.all(res[i0, i1, i2, i3] == a))
            assert_(res.shape == self.tgtshape[(4 - ip, 4 - jp)])
            assert_(not res.flags['OWNDATA'])


class TestMoveaxis:
    def test_move_to_end(self):
        x = np.random.randn(5, 6, 7)
        for source, expected in [(0, (6, 7, 5)),
                                 (1, (5, 7, 6)),
                                 (2, (5, 6, 7)),
                                 (-1, (5, 6, 7))]:
            actual = np.moveaxis(x, source, -1).shape
            assert_(actual, expected)

    def test_move_new_position(self):
        x = np.random.randn(1, 2, 3, 4)
        for source, destination, expected in [
                (0, 1, (2, 1, 3, 4)),
                (1, 2, (1, 3, 2, 4)),
                (1, -1, (1, 3, 4, 2)),
                ]:
            actual = np.moveaxis(x, source, destination).shape
            assert_(actual, expected)

    def test_preserve_order(self):
        x = np.zeros((1, 2, 3, 4))
        for source, destination in [
                (0, 0),
                (3, -1),
                (-1, 3),
                ([0, -1], [0, -1]),
                ([2, 0], [2, 0]),
                (range(4), range(4)),
                ]:
            actual = np.moveaxis(x, source, destination).shape
            assert_(actual, (1, 2, 3, 4))

    def test_move_multiples(self):
        x = np.zeros((0, 1, 2, 3))
        for source, destination, expected in [
                ([0, 1], [2, 3], (2, 3, 0, 1)),
                ([2, 3], [0, 1], (2, 3, 0, 1)),
                ([0, 1, 2], [2, 3, 0], (2, 3, 0, 1)),
                ([3, 0], [1, 0], (0, 3, 1, 2)),
                ([0, 3], [0, 1], (0, 3, 1, 2)),
                ]:
            actual = np.moveaxis(x, source, destination).shape
            assert_(actual, expected)

    def test_errors(self):
        x = np.random.randn(1, 2, 3)
        assert_raises_regex(AxisError, 'source.*out of bounds',
                            np.moveaxis, x, 3, 0)
        assert_raises_regex(AxisError, 'source.*out of bounds',
                            np.moveaxis, x, -4, 0)
        assert_raises_regex(AxisError, 'destination.*out of bounds',
                            np.moveaxis, x, 0, 5)
        assert_raises_regex(ValueError, 'repeated axis in `source`',
                            np.moveaxis, x, [0, 0], [0, 1])
        assert_raises_regex(ValueError, 'repeated axis in `destination`',
                            np.moveaxis, x, [0, 1], [1, 1])
        assert_raises_regex(ValueError, 'must have the same number',
                            np.moveaxis, x, 0, [0, 1])
        assert_raises_regex(ValueError, 'must have the same number',
                            np.moveaxis, x, [0, 1], [0])

    def test_array_likes(self):
        x = np.ma.zeros((1, 2, 3))
        result = np.moveaxis(x, 0, 0)
        assert_(x.shape, result.shape)
        assert_(isinstance(result, np.ma.MaskedArray))

        x = [1, 2, 3]
        result = np.moveaxis(x, 0, 0)
        assert_(x, list(result))
        assert_(isinstance(result, np.ndarray))


class TestCross:
    @pytest.mark.filterwarnings(
        "ignore:.*2-dimensional vectors.*:DeprecationWarning"
    )
    def test_2x2(self):
        u = [1, 2]
        v = [3, 4]
        z = -2
        cp = np.cross(u, v)
        assert_equal(cp, z)
        cp = np.cross(v, u)
        assert_equal(cp, -z)

    @pytest.mark.filterwarnings(
        "ignore:.*2-dimensional vectors.*:DeprecationWarning"
    )
    def test_2x3(self):
        u = [1, 2]
        v = [3, 4, 5]
        z = np.array([10, -5, -2])
        cp = np.cross(u, v)
        assert_equal(cp, z)
        cp = np.cross(v, u)
        assert_equal(cp, -z)

    def test_3x3(self):
        u = [1, 2, 3]
        v = [4, 5, 6]
        z = np.array([-3, 6, -3])
        cp = np.cross(u, v)
        assert_equal(cp, z)
        cp = np.cross(v, u)
        assert_equal(cp, -z)

    @pytest.mark.filterwarnings(
        "ignore:.*2-dimensional vectors.*:DeprecationWarning"
    )
    def test_broadcasting(self):
        # Ticket #2624 (Trac #2032)
        u = np.tile([1, 2], (11, 1))
        v = np.tile([3, 4], (11, 1))
        z = -2
        assert_equal(np.cross(u, v), z)
        assert_equal(np.cross(v, u), -z)
        assert_equal(np.cross(u, u), 0)

        u = np.tile([1, 2], (11, 1)).T
        v = np.tile([3, 4, 5], (11, 1))
        z = np.tile([10, -5, -2], (11, 1))
        assert_equal(np.cross(u, v, axisa=0), z)
        assert_equal(np.cross(v, u.T), -z)
        assert_equal(np.cross(v, v), 0)

        u = np.tile([1, 2, 3], (11, 1)).T
        v = np.tile([3, 4], (11, 1)).T
        z = np.tile([-12, 9, -2], (11, 1))
        assert_equal(np.cross(u, v, axisa=0, axisb=0), z)
        assert_equal(np.cross(v.T, u.T), -z)
        assert_equal(np.cross(u.T, u.T), 0)

        u = np.tile([1, 2, 3], (5, 1))
        v = np.tile([4, 5, 6], (5, 1)).T
        z = np.tile([-3, 6, -3], (5, 1))
        assert_equal(np.cross(u, v, axisb=0), z)
        assert_equal(np.cross(v.T, u), -z)
        assert_equal(np.cross(u, u), 0)

    @pytest.mark.filterwarnings(
        "ignore:.*2-dimensional vectors.*:DeprecationWarning"
    )
    def test_broadcasting_shapes(self):
        u = np.ones((2, 1, 3))
        v = np.ones((5, 3))
        assert_equal(np.cross(u, v).shape, (2, 5, 3))
        u = np.ones((10, 3, 5))
        v = np.ones((2, 5))
        assert_equal(np.cross(u, v, axisa=1, axisb=0).shape, (10, 5, 3))
        assert_raises(AxisError, np.cross, u, v, axisa=1, axisb=2)
        assert_raises(AxisError, np.cross, u, v, axisa=3, axisb=0)
        u = np.ones((10, 3, 5, 7))
        v = np.ones((5, 7, 2))
        assert_equal(np.cross(u, v, axisa=1, axisc=2).shape, (10, 5, 3, 7))
        assert_raises(AxisError, np.cross, u, v, axisa=-5, axisb=2)
        assert_raises(AxisError, np.cross, u, v, axisa=1, axisb=-4)
        # gh-5885
        u = np.ones((3, 4, 2))
        for axisc in range(-2, 2):
            assert_equal(np.cross(u, u, axisc=axisc).shape, (3, 4))

    def test_uint8_int32_mixed_dtypes(self):
        # regression test for gh-19138
        u = np.array([[195, 8, 9]], np.uint8)
        v = np.array([250, 166, 68], np.int32)
        z = np.array([[950, 11010, -30370]], dtype=np.int32)
        assert_equal(np.cross(v, u), z)
        assert_equal(np.cross(u, v), -z)

    @pytest.mark.parametrize("a, b", [(0, [1, 2]), ([1, 2], 3)])
    def test_zero_dimension(self, a, b):
        with pytest.raises(ValueError) as exc:
            np.cross(a, b)
        assert "At least one array has zero dimension" in str(exc.value)


def test_outer_out_param():
    arr1 = np.ones((5,))
    arr2 = np.ones((2,))
    arr3 = np.linspace(-2, 2, 5)
    out1 = np.ndarray(shape=(5, 5))
    out2 = np.ndarray(shape=(2, 5))
    res1 = np.outer(arr1, arr3, out1)
    assert_equal(res1, out1)
    assert_equal(np.outer(arr2, arr3, out2), out2)


class TestIndices:

    def test_simple(self):
        [x, y] = np.indices((4, 3))
        assert_array_equal(x, np.array([[0, 0, 0],
                                        [1, 1, 1],
                                        [2, 2, 2],
                                        [3, 3, 3]]))
        assert_array_equal(y, np.array([[0, 1, 2],
                                        [0, 1, 2],
                                        [0, 1, 2],
                                        [0, 1, 2]]))

    def test_single_input(self):
        [x] = np.indices((4,))
        assert_array_equal(x, np.array([0, 1, 2, 3]))

        [x] = np.indices((4,), sparse=True)
        assert_array_equal(x, np.array([0, 1, 2, 3]))

    def test_scalar_input(self):
        assert_array_equal([], np.indices(()))
        assert_array_equal([], np.indices((), sparse=True))
        assert_array_equal([[]], np.indices((0,)))
        assert_array_equal([[]], np.indices((0,), sparse=True))

    def test_sparse(self):
        [x, y] = np.indices((4, 3), sparse=True)
        assert_array_equal(x, np.array([[0], [1], [2], [3]]))
        assert_array_equal(y, np.array([[0, 1, 2]]))

    @pytest.mark.parametrize("dtype", [np.int32, np.int64, np.float32, np.float64])
    @pytest.mark.parametrize("dims", [(), (0,), (4, 3)])
    def test_return_type(self, dtype, dims):
        inds = np.indices(dims, dtype=dtype)
        assert_(inds.dtype == dtype)

        for arr in np.indices(dims, dtype=dtype, sparse=True):
            assert_(arr.dtype == dtype)


class TestRequire:
    flag_names = ['C', 'C_CONTIGUOUS', 'CONTIGUOUS',
                  'F', 'F_CONTIGUOUS', 'FORTRAN',
                  'A', 'ALIGNED',
                  'W', 'WRITEABLE',
                  'O', 'OWNDATA']

    def generate_all_false(self, dtype):
        arr = np.zeros((2, 2), [('junk', 'i1'), ('a', dtype)])
        arr.setflags(write=False)
        a = arr['a']
        assert_(not a.flags['C'])
        assert_(not a.flags['F'])
        assert_(not a.flags['O'])
        assert_(not a.flags['W'])
        assert_(not a.flags['A'])
        return a

    def set_and_check_flag(self, flag, dtype, arr):
        if dtype is None:
            dtype = arr.dtype
        b = np.require(arr, dtype, [flag])
        assert_(b.flags[flag])
        assert_(b.dtype == dtype)

        # a further call to np.require ought to return the same array
        # unless OWNDATA is specified.
        c = np.require(b, None, [flag])
        if flag[0] != 'O':
            assert_(c is b)
        else:
            assert_(c.flags[flag])

    def test_require_each(self):

        id = ['f8', 'i4']
        fd = [None, 'f8', 'c16']
        for idtype, fdtype, flag in itertools.product(id, fd, self.flag_names):
            a = self.generate_all_false(idtype)
            self.set_and_check_flag(flag, fdtype,  a)

    def test_unknown_requirement(self):
        a = self.generate_all_false('f8')
        assert_raises(KeyError, np.require, a, None, 'Q')

    def test_non_array_input(self):
        a = np.require([1, 2, 3, 4], 'i4', ['C', 'A', 'O'])
        assert_(a.flags['O'])
        assert_(a.flags['C'])
        assert_(a.flags['A'])
        assert_(a.dtype == 'i4')
        assert_equal(a, [1, 2, 3, 4])

    def test_C_and_F_simul(self):
        a = self.generate_all_false('f8')
        assert_raises(ValueError, np.require, a, None, ['C', 'F'])

    def test_ensure_array(self):
        class ArraySubclass(np.ndarray):
            pass

        a = ArraySubclass((2, 2))
        b = np.require(a, None, ['E'])
        assert_(type(b) is np.ndarray)

    def test_preserve_subtype(self):
        class ArraySubclass(np.ndarray):
            pass

        for flag in self.flag_names:
            a = ArraySubclass((2, 2))
            self.set_and_check_flag(flag, None, a)


class TestBroadcast:
    def test_broadcast_in_args(self):
        # gh-5881
        arrs = [np.empty((6, 7)), np.empty((5, 6, 1)), np.empty((7,)),
                np.empty((5, 1, 7))]
        mits = [np.broadcast(*arrs),
                np.broadcast(np.broadcast(*arrs[:0]), np.broadcast(*arrs[0:])),
                np.broadcast(np.broadcast(*arrs[:1]), np.broadcast(*arrs[1:])),
                np.broadcast(np.broadcast(*arrs[:2]), np.broadcast(*arrs[2:])),
                np.broadcast(arrs[0], np.broadcast(*arrs[1:-1]), arrs[-1])]
        for mit in mits:
            assert_equal(mit.shape, (5, 6, 7))
            assert_equal(mit.ndim, 3)
            assert_equal(mit.nd, 3)
            assert_equal(mit.numiter, 4)
            for a, ia in zip(arrs, mit.iters):
                assert_(a is ia.base)

    def test_broadcast_single_arg(self):
        # gh-6899
        arrs = [np.empty((5, 6, 7))]
        mit = np.broadcast(*arrs)
        assert_equal(mit.shape, (5, 6, 7))
        assert_equal(mit.ndim, 3)
        assert_equal(mit.nd, 3)
        assert_equal(mit.numiter, 1)
        assert_(arrs[0] is mit.iters[0].base)

    def test_number_of_arguments(self):
        arr = np.empty((5,))
        for j in range(70):
            arrs = [arr] * j
            if j > 64:
                assert_raises(ValueError, np.broadcast, *arrs)
            else:
                mit = np.broadcast(*arrs)
                assert_equal(mit.numiter, j)

    def test_broadcast_error_kwargs(self):
        # gh-13455
        arrs = [np.empty((5, 6, 7))]
        mit = np.broadcast(*arrs)
        mit2 = np.broadcast(*arrs, **{})  # noqa: PIE804
        assert_equal(mit.shape, mit2.shape)
        assert_equal(mit.ndim, mit2.ndim)
        assert_equal(mit.nd, mit2.nd)
        assert_equal(mit.numiter, mit2.numiter)
        assert_(mit.iters[0].base is mit2.iters[0].base)

        assert_raises(ValueError, np.broadcast, 1, x=1)

    def test_shape_mismatch_error_message(self):
        with pytest.raises(ValueError, match=r"arg 0 with shape \(1, 3\) and "
                                             r"arg 2 with shape \(2,\)"):
            np.broadcast([[1, 2, 3]], [[4], [5]], [6, 7])


class TestKeepdims:

    class sub_array(np.ndarray):
        def sum(self, axis=None, dtype=None, out=None):
            return np.ndarray.sum(self, axis, dtype, out, keepdims=True)

    def test_raise(self):
        sub_class = self.sub_array
        x = np.arange(30).view(sub_class)
        assert_raises(TypeError, np.sum, x, keepdims=True)


class TestTensordot:

    def test_zero_dimension(self):
        # Test resolution to issue #5663
        a = np.ndarray((3, 0))
        b = np.ndarray((0, 4))
        td = np.tensordot(a, b, (1, 0))
        assert_array_equal(td, np.dot(a, b))
        assert_array_equal(td, np.einsum('ij,jk', a, b))

    def test_zero_dimensional(self):
        # gh-12130
        arr_0d = np.array(1)
        ret = np.tensordot(arr_0d, arr_0d, ([], []))  # contracting no axes is well defined
        assert_array_equal(ret, arr_0d)


class TestAsType:

    def test_astype(self):
        data = [[1, 2], [3, 4]]
        actual = np.astype(
            np.array(data, dtype=np.int64), np.uint32
        )
        expected = np.array(data, dtype=np.uint32)

        assert_array_equal(actual, expected)
        assert_equal(actual.dtype, expected.dtype)

        assert np.shares_memory(
            actual, np.astype(actual, actual.dtype, copy=False)
        )

        actual = np.astype(np.int64(10), np.float64)
        expected = np.float64(10)
        assert_equal(actual, expected)
        assert_equal(actual.dtype, expected.dtype)

        with pytest.raises(TypeError, match="Input should be a NumPy array"):
            np.astype(data, np.float64)