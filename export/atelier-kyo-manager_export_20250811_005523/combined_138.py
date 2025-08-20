
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\profile_result_processor.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation. All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

"""This profiler result processor print out the kernel time spent on each Node of the model.
Example of importing profile result file from onnxruntime_perf_test:
    python profile_result_processor.py --input profile_2021-10-25_12-02-41.json
"""

import argparse
import json

_NODES_TYPE_CONTAINING_SUBGRAPH = frozenset(("Scan", "Loop", "If"))


def parse_arguments(argv=None):
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "-i",
        "--input",
        required=False,
        type=str,
        help="Set the input file for reading the profile results",
    )

    parser.add_argument(
        "--threshold",
        required=False,
        type=float,
        default=0.01,
        help="Threshold of run time ratio among all nodes. Nodes with larger ratio will show in top expensive nodes.",
    )

    parser.add_argument(
        "--provider",
        required=False,
        type=str,
        default="cuda",
        help="Execution provider to use",
    )

    parser.add_argument(
        "--kernel_time_only",
        required=False,
        action="store_true",
        help="Only include the kernel time and no fence time",
    )

    parser.set_defaults(kernel_time_only=False)

    parser.add_argument("-v", "--verbose", required=False, action="store_true")
    parser.set_defaults(verbose=False)

    return parser.parse_args(argv)


def load_profile_json(profile_file):
    print(f"loading profile output {profile_file} ...")

    with open(profile_file) as opened_file:
        sess_time = json.load(opened_file)

    assert isinstance(sess_time, list)
    return sess_time


def parse_kernel_results(sess_time, threshold=0):
    """Parse profile data and output nodes in two sections - nodes in the original order, and top expensive nodes.

    Args:
        sess_time (List[Dict]): profile data
        threshold (int, optional): Minimum ratio of duration among all. Defaults to 0.

    Returns:
        List[str]: lines of string for output.
    """
    kernel_name_to_op_name = {}
    kernel_time = {}
    kernel_freq = {}
    total = 0
    session_init = False
    for item in sess_time:
        # Skip all MemcpyHostToDevice before session_initialization
        if item["cat"] == "Session" and item["name"] == "session_initialization":
            session_init = True
        if not session_init:
            continue

        if item["cat"] == "Kernel" and "dur" in item and "args" in item and "op_name" in item["args"]:
            kernel_name = item["name"]

            op_name = item["args"]["op_name"]
            if op_name in _NODES_TYPE_CONTAINING_SUBGRAPH:
                continue

            # Handle MemcpyHostToDevice and MemcpyDeviceToHost here
            if not op_name:
                op_name = f"({kernel_name})"

            if kernel_name in kernel_time:
                kernel_time[kernel_name] += item["dur"]
                kernel_freq[kernel_name] += 1
            else:
                kernel_time[kernel_name] = item["dur"]
                kernel_freq[kernel_name] = 1
                kernel_name_to_op_name[kernel_name] = op_name

            total += item["dur"]

    if not kernel_time:
        return ["No kernel record found!"]

    # Output items with run time ratio > thresholds, and sorted by duration in the descending order.
    lines = []
    lines.append(f"\nTop expensive kernels with Time% >= {threshold * 100:.2f}:")
    lines.append("-" * 64)
    lines.append("Total(μs)\tTime%\tCalls\tAvg(μs)\tKernel")
    for kernel_name, duration in sorted(kernel_time.items(), key=lambda x: x[1], reverse=True):
        ratio = duration / total
        if ratio < threshold:
            continue

        calls = kernel_freq[kernel_name]
        avg_time = duration / float(calls)
        lines.append(f"{duration:10d}\t{ratio * 100.0:5.2f}\t{calls:5d}\t{avg_time:8.1f}\t{kernel_name}")

    # Group by operator
    op_time = {}
    for kernel_name, op_name in kernel_name_to_op_name.items():
        duration = kernel_time[kernel_name]
        if op_name in op_time:
            op_time[op_name] += duration
        else:
            op_time[op_name] = duration

    lines.append("\nGroup kernel time by operator:")
    lines.append("-" * 64)
    lines.append("Total(μs)\tTime%\tOperator")
    for op_name, duration in sorted(op_time.items(), key=lambda x: x[1], reverse=True):
        ratio = duration / total
        lines.append(f"{duration:10d}\t{ratio * 100.0:5.2f}\t{op_name}")

    return lines


def parse_node_results(sess_time, kernel_time_only=False, threshold=0):
    """Parse profile data and output nodes in two sections - nodes in the original order, and top expensive nodes.

    Args:
        sess_time (List[Dict]): profile data
        kernel_time_only (bool, optional): Only include items for kernel time. Defaults to False.
        threshold (int, optional): Minimum ratio of duration among all. Defaults to 0.

    Returns:
        List[str]: lines of string for output.
    """
    node_name_list = []
    node_time = {}
    node_freq = {}
    node_provider = {}
    total = 0
    for item in sess_time:
        if item["cat"] == "Node" and "dur" in item and "args" in item and "op_name" in item["args"]:
            node_name = (
                item["name"].replace("_kernel_time", "").replace("_fence_before", "").replace("_fence_after", "")
            )

            if "provider" in item["args"]:
                if item["args"]["provider"] == "CPUExecutionProvider":
                    device = "CPU"
                elif item["args"]["provider"] == "CUDAExecutionProvider":
                    device = "CUDA"
                elif item["args"]["provider"] == "DmlExecutionProvider":
                    device = "DML"

                if node_name not in node_provider:
                    node_provider[node_name] = device
                else:
                    assert node_provider[node_name] == device
            elif kernel_time_only:
                continue

            op_name = item["args"]["op_name"]
            if op_name in _NODES_TYPE_CONTAINING_SUBGRAPH:
                continue

            if node_name in node_time:
                node_time[node_name] += item["dur"]
                node_freq[node_name] += 1
            else:
                node_time[node_name] = item["dur"]
                node_freq[node_name] = 1
                node_name_list.append(node_name)

            total += item["dur"]

    # Output items in the original order.
    lines = [
        "\nNodes in the original order:",
        "-" * 64,
        "Total(μs)\tTime%\tAcc %\tAvg(μs)\tCalls\tProvider\tNode",
    ]
    before_percentage = 0.0
    for node_name in node_name_list:
        duration = node_time[node_name]
        calls = node_freq[node_name]
        avg_time = duration / float(calls)
        percentage = (duration / total) * 100.0
        provider = node_provider.get(node_name, "")
        before_percentage += percentage
        lines.append(
            f"{duration:10d}\t{percentage:5.2f}\t{before_percentage:5.2f}\t{avg_time:8.1f}\t{calls:5d}\t{provider:8s}\t{node_name}"
        )

    # Output items with run time ratio > thresholds, and sorted by duration in the descending order.
    lines.append(f"\nTop expensive nodes with Time% >= {threshold * 100:.2f}:")
    lines.append("-" * 64)
    lines.append("Total(μs)\tTime%\tAvg(μs)\tCalls\tProvider\tNode")
    for node_name, duration in sorted(node_time.items(), key=lambda x: x[1], reverse=True):
        ratio = duration / total
        if ratio < threshold:
            continue

        calls = node_freq[node_name]
        avg_time = duration / float(calls)
        percentage = (duration / total) * 100.0
        provider = node_provider.get(node_name, "")
        lines.append(f"{duration:10d}\t{percentage:5.2f}\t{avg_time:8.1f}\t{calls:5d}\t{provider:8s}\t{node_name}")

    return lines


def group_node_results(sess_time):
    """Group results by operator name.

    Args:
        sess_time (List[Dict]): profile data

    Returns:
        List[str]: lines of string for output.
    """
    op_kernel_time = {}
    op_kernel_records = {}
    total_kernel_time = 0

    provider_op_kernel_time = {}
    provider_op_kernel_records = {}
    provider_kernel_time = {}

    op_fence_time = {}
    total_fence_time = 0

    provider_counter = {}
    for item in sess_time:
        if item["cat"] == "Node" and "dur" in item and "args" in item and "op_name" in item["args"]:
            op_name = item["args"]["op_name"]

            # TODO: shall we have a separated group for nodes with subgraph?
            if op_name in _NODES_TYPE_CONTAINING_SUBGRAPH:
                continue

            if "provider" not in item["args"]:
                if "fence" in item["name"]:
                    if op_name in op_fence_time:
                        op_fence_time[op_name] += item["dur"]
                    else:
                        op_fence_time[op_name] = item["dur"]
                    total_fence_time += item["dur"]
                continue

            provider = item["args"].get("provider", "")
            if provider in provider_counter:
                provider_counter[provider] += 1
            else:
                provider_counter[provider] = 1

            key = f"{provider}:{op_name}"
            if key in provider_op_kernel_time:
                provider_op_kernel_time[key] += item["dur"]
                provider_op_kernel_records[key] += 1
            else:
                provider_op_kernel_time[key] = item["dur"]
                provider_op_kernel_records[key] = 1

            if provider in provider_kernel_time:
                provider_kernel_time[provider] += item["dur"]
            else:
                provider_kernel_time[provider] = item["dur"]

            if op_name in op_kernel_time:
                op_kernel_time[op_name] += item["dur"]
                op_kernel_records[op_name] += 1
            else:
                op_kernel_time[op_name] = item["dur"]
                op_kernel_records[op_name] = 1

            total_kernel_time += item["dur"]

    lines = ["", "Grouped by operator"]
    lines.append("-" * 64)
    lines.append("Total(μs)\tTime%\tKernel(μs)\tKernel%\tCalls\tAvgKernel(μs)\tFence(μs)\tOperator")
    for op_name, kernel_time in sorted(op_kernel_time.items(), key=lambda x: x[1], reverse=True):
        fence_time = op_fence_time.get(op_name, 0)
        kernel_time_ratio = kernel_time / total_kernel_time
        total_time = kernel_time + fence_time
        time_ratio = total_time / (total_kernel_time + total_fence_time)
        kernel_calls = op_kernel_records[op_name]
        avg_kernel_time = kernel_time / kernel_calls
        lines.append(
            f"{total_time:10d}\t{time_ratio * 100.0:5.2f}\t{kernel_time:11d}\t{kernel_time_ratio * 100.0:5.2f}\t{kernel_calls:5d}\t{avg_kernel_time:14.1f}\t{fence_time:10d}\t{op_name}"
        )

    lines += ["", "Grouped by provider + operator"]
    lines.append("-" * 64)
    lines.append("Kernel(μs)\tProvider%\tCalls\tAvgKernel(μs)\tProvider\tOperator")
    for key, kernel_time in sorted(provider_op_kernel_time.items(), key=lambda x: x[1], reverse=True):
        parts = key.split(":")
        provider = parts[0]
        op_name = parts[1]
        short_ep = provider.replace("ExecutionProvider", "")
        calls = provider_op_kernel_records[key]
        avg_kernel_time = kernel_time / calls
        provider_time_ratio = kernel_time / provider_kernel_time[provider]
        lines.append(
            f"{kernel_time:10d}\t{provider_time_ratio * 100.0:9.2f}\t{calls:5d}\t{avg_kernel_time:14.1f}\t{short_ep:8s}\t{op_name}"
        )

    return lines


def process_results(profile_file, args):
    profile_records = load_profile_json(profile_file)

    lines = parse_kernel_results(profile_records, args.threshold)

    lines += parse_node_results(profile_records, args.kernel_time_only, args.threshold)

    lines += group_node_results(profile_records)

    return lines


if __name__ == "__main__":
    arguments = parse_arguments()
    print("Arguments", arguments)

    from benchmark_helper import setup_logger

    setup_logger(arguments.verbose)

    profile_file = arguments.input

    results = process_results(profile_file, arguments)

    for line in results:
        print(line)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\distlib\resources.py ===
# -*- coding: utf-8 -*-
#
# Copyright (C) 2013-2017 Vinay Sajip.
# Licensed to the Python Software Foundation under a contributor agreement.
# See LICENSE.txt and CONTRIBUTORS.txt.
#
from __future__ import unicode_literals

import bisect
import io
import logging
import os
import pkgutil
import sys
import types
import zipimport

from . import DistlibException
from .util import cached_property, get_cache_base, Cache

logger = logging.getLogger(__name__)


cache = None    # created when needed


class ResourceCache(Cache):
    def __init__(self, base=None):
        if base is None:
            # Use native string to avoid issues on 2.x: see Python #20140.
            base = os.path.join(get_cache_base(), str('resource-cache'))
        super(ResourceCache, self).__init__(base)

    def is_stale(self, resource, path):
        """
        Is the cache stale for the given resource?

        :param resource: The :class:`Resource` being cached.
        :param path: The path of the resource in the cache.
        :return: True if the cache is stale.
        """
        # Cache invalidation is a hard problem :-)
        return True

    def get(self, resource):
        """
        Get a resource into the cache,

        :param resource: A :class:`Resource` instance.
        :return: The pathname of the resource in the cache.
        """
        prefix, path = resource.finder.get_cache_info(resource)
        if prefix is None:
            result = path
        else:
            result = os.path.join(self.base, self.prefix_to_dir(prefix), path)
            dirname = os.path.dirname(result)
            if not os.path.isdir(dirname):
                os.makedirs(dirname)
            if not os.path.exists(result):
                stale = True
            else:
                stale = self.is_stale(resource, path)
            if stale:
                # write the bytes of the resource to the cache location
                with open(result, 'wb') as f:
                    f.write(resource.bytes)
        return result


class ResourceBase(object):
    def __init__(self, finder, name):
        self.finder = finder
        self.name = name


class Resource(ResourceBase):
    """
    A class representing an in-package resource, such as a data file. This is
    not normally instantiated by user code, but rather by a
    :class:`ResourceFinder` which manages the resource.
    """
    is_container = False        # Backwards compatibility

    def as_stream(self):
        """
        Get the resource as a stream.

        This is not a property to make it obvious that it returns a new stream
        each time.
        """
        return self.finder.get_stream(self)

    @cached_property
    def file_path(self):
        global cache
        if cache is None:
            cache = ResourceCache()
        return cache.get(self)

    @cached_property
    def bytes(self):
        return self.finder.get_bytes(self)

    @cached_property
    def size(self):
        return self.finder.get_size(self)


class ResourceContainer(ResourceBase):
    is_container = True     # Backwards compatibility

    @cached_property
    def resources(self):
        return self.finder.get_resources(self)


class ResourceFinder(object):
    """
    Resource finder for file system resources.
    """

    if sys.platform.startswith('java'):
        skipped_extensions = ('.pyc', '.pyo', '.class')
    else:
        skipped_extensions = ('.pyc', '.pyo')

    def __init__(self, module):
        self.module = module
        self.loader = getattr(module, '__loader__', None)
        self.base = os.path.dirname(getattr(module, '__file__', ''))

    def _adjust_path(self, path):
        return os.path.realpath(path)

    def _make_path(self, resource_name):
        # Issue #50: need to preserve type of path on Python 2.x
        # like os.path._get_sep
        if isinstance(resource_name, bytes):    # should only happen on 2.x
            sep = b'/'
        else:
            sep = '/'
        parts = resource_name.split(sep)
        parts.insert(0, self.base)
        result = os.path.join(*parts)
        return self._adjust_path(result)

    def _find(self, path):
        return os.path.exists(path)

    def get_cache_info(self, resource):
        return None, resource.path

    def find(self, resource_name):
        path = self._make_path(resource_name)
        if not self._find(path):
            result = None
        else:
            if self._is_directory(path):
                result = ResourceContainer(self, resource_name)
            else:
                result = Resource(self, resource_name)
            result.path = path
        return result

    def get_stream(self, resource):
        return open(resource.path, 'rb')

    def get_bytes(self, resource):
        with open(resource.path, 'rb') as f:
            return f.read()

    def get_size(self, resource):
        return os.path.getsize(resource.path)

    def get_resources(self, resource):
        def allowed(f):
            return (f != '__pycache__' and not
                    f.endswith(self.skipped_extensions))
        return set([f for f in os.listdir(resource.path) if allowed(f)])

    def is_container(self, resource):
        return self._is_directory(resource.path)

    _is_directory = staticmethod(os.path.isdir)

    def iterator(self, resource_name):
        resource = self.find(resource_name)
        if resource is not None:
            todo = [resource]
            while todo:
                resource = todo.pop(0)
                yield resource
                if resource.is_container:
                    rname = resource.name
                    for name in resource.resources:
                        if not rname:
                            new_name = name
                        else:
                            new_name = '/'.join([rname, name])
                        child = self.find(new_name)
                        if child.is_container:
                            todo.append(child)
                        else:
                            yield child


class ZipResourceFinder(ResourceFinder):
    """
    Resource finder for resources in .zip files.
    """
    def __init__(self, module):
        super(ZipResourceFinder, self).__init__(module)
        archive = self.loader.archive
        self.prefix_len = 1 + len(archive)
        # PyPy doesn't have a _files attr on zipimporter, and you can't set one
        if hasattr(self.loader, '_files'):
            self._files = self.loader._files
        else:
            self._files = zipimport._zip_directory_cache[archive]
        self.index = sorted(self._files)

    def _adjust_path(self, path):
        return path

    def _find(self, path):
        path = path[self.prefix_len:]
        if path in self._files:
            result = True
        else:
            if path and path[-1] != os.sep:
                path = path + os.sep
            i = bisect.bisect(self.index, path)
            try:
                result = self.index[i].startswith(path)
            except IndexError:
                result = False
        if not result:
            logger.debug('_find failed: %r %r', path, self.loader.prefix)
        else:
            logger.debug('_find worked: %r %r', path, self.loader.prefix)
        return result

    def get_cache_info(self, resource):
        prefix = self.loader.archive
        path = resource.path[1 + len(prefix):]
        return prefix, path

    def get_bytes(self, resource):
        return self.loader.get_data(resource.path)

    def get_stream(self, resource):
        return io.BytesIO(self.get_bytes(resource))

    def get_size(self, resource):
        path = resource.path[self.prefix_len:]
        return self._files[path][3]

    def get_resources(self, resource):
        path = resource.path[self.prefix_len:]
        if path and path[-1] != os.sep:
            path += os.sep
        plen = len(path)
        result = set()
        i = bisect.bisect(self.index, path)
        while i < len(self.index):
            if not self.index[i].startswith(path):
                break
            s = self.index[i][plen:]
            result.add(s.split(os.sep, 1)[0])   # only immediate children
            i += 1
        return result

    def _is_directory(self, path):
        path = path[self.prefix_len:]
        if path and path[-1] != os.sep:
            path += os.sep
        i = bisect.bisect(self.index, path)
        try:
            result = self.index[i].startswith(path)
        except IndexError:
            result = False
        return result


_finder_registry = {
    type(None): ResourceFinder,
    zipimport.zipimporter: ZipResourceFinder
}

try:
    # In Python 3.6, _frozen_importlib -> _frozen_importlib_external
    try:
        import _frozen_importlib_external as _fi
    except ImportError:
        import _frozen_importlib as _fi
    _finder_registry[_fi.SourceFileLoader] = ResourceFinder
    _finder_registry[_fi.FileFinder] = ResourceFinder
    # See issue #146
    _finder_registry[_fi.SourcelessFileLoader] = ResourceFinder
    del _fi
except (ImportError, AttributeError):
    pass


def register_finder(loader, finder_maker):
    _finder_registry[type(loader)] = finder_maker


_finder_cache = {}


def finder(package):
    """
    Return a resource finder for a package.
    :param package: The name of the package.
    :return: A :class:`ResourceFinder` instance for the package.
    """
    if package in _finder_cache:
        result = _finder_cache[package]
    else:
        if package not in sys.modules:
            __import__(package)
        module = sys.modules[package]
        path = getattr(module, '__path__', None)
        if path is None:
            raise DistlibException('You cannot get a finder for a module, '
                                   'only for a package')
        loader = getattr(module, '__loader__', None)
        finder_maker = _finder_registry.get(type(loader))
        if finder_maker is None:
            raise DistlibException('Unable to locate finder for %r' % package)
        result = finder_maker(module)
        _finder_cache[package] = result
    return result


_dummy_module = types.ModuleType(str('__dummy__'))


def finder_for_path(path):
    """
    Return a resource finder for a path, which should represent a container.

    :param path: The path.
    :return: A :class:`ResourceFinder` instance for the path.
    """
    result = None
    # calls any path hooks, gets importer into cache
    pkgutil.get_importer(path)
    loader = sys.path_importer_cache.get(path)
    finder = _finder_registry.get(type(loader))
    if finder:
        module = _dummy_module
        module.__file__ = os.path.join(path, '')
        module.__loader__ = loader
        result = finder(module)
    return result

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\io\excel\_odswriter.py ===
from __future__ import annotations

from collections import defaultdict
import datetime
import json
from typing import (
    TYPE_CHECKING,
    Any,
    DefaultDict,
    cast,
    overload,
)

from pandas.io.excel._base import ExcelWriter
from pandas.io.excel._util import (
    combine_kwargs,
    validate_freeze_panes,
)

if TYPE_CHECKING:
    from pandas._typing import (
        ExcelWriterIfSheetExists,
        FilePath,
        StorageOptions,
        WriteExcelBuffer,
    )

    from pandas.io.formats.excel import ExcelCell


class ODSWriter(ExcelWriter):
    _engine = "odf"
    _supported_extensions = (".ods",)

    def __init__(
        self,
        path: FilePath | WriteExcelBuffer | ExcelWriter,
        engine: str | None = None,
        date_format: str | None = None,
        datetime_format=None,
        mode: str = "w",
        storage_options: StorageOptions | None = None,
        if_sheet_exists: ExcelWriterIfSheetExists | None = None,
        engine_kwargs: dict[str, Any] | None = None,
        **kwargs,
    ) -> None:
        from odf.opendocument import OpenDocumentSpreadsheet

        if mode == "a":
            raise ValueError("Append mode is not supported with odf!")

        engine_kwargs = combine_kwargs(engine_kwargs, kwargs)
        self._book = OpenDocumentSpreadsheet(**engine_kwargs)

        super().__init__(
            path,
            mode=mode,
            storage_options=storage_options,
            if_sheet_exists=if_sheet_exists,
            engine_kwargs=engine_kwargs,
        )

        self._style_dict: dict[str, str] = {}

    @property
    def book(self):
        """
        Book instance of class odf.opendocument.OpenDocumentSpreadsheet.

        This attribute can be used to access engine-specific features.
        """
        return self._book

    @property
    def sheets(self) -> dict[str, Any]:
        """Mapping of sheet names to sheet objects."""
        from odf.table import Table

        result = {
            sheet.getAttribute("name"): sheet
            for sheet in self.book.getElementsByType(Table)
        }
        return result

    def _save(self) -> None:
        """
        Save workbook to disk.
        """
        for sheet in self.sheets.values():
            self.book.spreadsheet.addElement(sheet)
        self.book.save(self._handles.handle)

    def _write_cells(
        self,
        cells: list[ExcelCell],
        sheet_name: str | None = None,
        startrow: int = 0,
        startcol: int = 0,
        freeze_panes: tuple[int, int] | None = None,
    ) -> None:
        """
        Write the frame cells using odf
        """
        from odf.table import (
            Table,
            TableCell,
            TableRow,
        )
        from odf.text import P

        sheet_name = self._get_sheet_name(sheet_name)
        assert sheet_name is not None

        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = Table(name=sheet_name)
            self.book.spreadsheet.addElement(wks)

        if validate_freeze_panes(freeze_panes):
            freeze_panes = cast(tuple[int, int], freeze_panes)
            self._create_freeze_panes(sheet_name, freeze_panes)

        for _ in range(startrow):
            wks.addElement(TableRow())

        rows: DefaultDict = defaultdict(TableRow)
        col_count: DefaultDict = defaultdict(int)

        for cell in sorted(cells, key=lambda cell: (cell.row, cell.col)):
            # only add empty cells if the row is still empty
            if not col_count[cell.row]:
                for _ in range(startcol):
                    rows[cell.row].addElement(TableCell())

            # fill with empty cells if needed
            for _ in range(cell.col - col_count[cell.row]):
                rows[cell.row].addElement(TableCell())
                col_count[cell.row] += 1

            pvalue, tc = self._make_table_cell(cell)
            rows[cell.row].addElement(tc)
            col_count[cell.row] += 1
            p = P(text=pvalue)
            tc.addElement(p)

        # add all rows to the sheet
        if len(rows) > 0:
            for row_nr in range(max(rows.keys()) + 1):
                wks.addElement(rows[row_nr])

    def _make_table_cell_attributes(self, cell) -> dict[str, int | str]:
        """Convert cell attributes to OpenDocument attributes

        Parameters
        ----------
        cell : ExcelCell
            Spreadsheet cell data

        Returns
        -------
        attributes : Dict[str, Union[int, str]]
            Dictionary with attributes and attribute values
        """
        attributes: dict[str, int | str] = {}
        style_name = self._process_style(cell.style)
        if style_name is not None:
            attributes["stylename"] = style_name
        if cell.mergestart is not None and cell.mergeend is not None:
            attributes["numberrowsspanned"] = max(1, cell.mergestart)
            attributes["numbercolumnsspanned"] = cell.mergeend
        return attributes

    def _make_table_cell(self, cell) -> tuple[object, Any]:
        """Convert cell data to an OpenDocument spreadsheet cell

        Parameters
        ----------
        cell : ExcelCell
            Spreadsheet cell data

        Returns
        -------
        pvalue, cell : Tuple[str, TableCell]
            Display value, Cell value
        """
        from odf.table import TableCell

        attributes = self._make_table_cell_attributes(cell)
        val, fmt = self._value_with_fmt(cell.val)
        pvalue = value = val
        if isinstance(val, bool):
            value = str(val).lower()
            pvalue = str(val).upper()
            return (
                pvalue,
                TableCell(
                    valuetype="boolean",
                    booleanvalue=value,
                    attributes=attributes,
                ),
            )
        elif isinstance(val, datetime.datetime):
            # Fast formatting
            value = val.isoformat()
            # Slow but locale-dependent
            pvalue = val.strftime("%c")
            return (
                pvalue,
                TableCell(valuetype="date", datevalue=value, attributes=attributes),
            )
        elif isinstance(val, datetime.date):
            # Fast formatting
            value = f"{val.year}-{val.month:02d}-{val.day:02d}"
            # Slow but locale-dependent
            pvalue = val.strftime("%x")
            return (
                pvalue,
                TableCell(valuetype="date", datevalue=value, attributes=attributes),
            )
        elif isinstance(val, str):
            return (
                pvalue,
                TableCell(
                    valuetype="string",
                    stringvalue=value,
                    attributes=attributes,
                ),
            )
        else:
            return (
                pvalue,
                TableCell(
                    valuetype="float",
                    value=value,
                    attributes=attributes,
                ),
            )

    @overload
    def _process_style(self, style: dict[str, Any]) -> str:
        ...

    @overload
    def _process_style(self, style: None) -> None:
        ...

    def _process_style(self, style: dict[str, Any] | None) -> str | None:
        """Convert a style dictionary to a OpenDocument style sheet

        Parameters
        ----------
        style : Dict
            Style dictionary

        Returns
        -------
        style_key : str
            Unique style key for later reference in sheet
        """
        from odf.style import (
            ParagraphProperties,
            Style,
            TableCellProperties,
            TextProperties,
        )

        if style is None:
            return None
        style_key = json.dumps(style)
        if style_key in self._style_dict:
            return self._style_dict[style_key]
        name = f"pd{len(self._style_dict)+1}"
        self._style_dict[style_key] = name
        odf_style = Style(name=name, family="table-cell")
        if "font" in style:
            font = style["font"]
            if font.get("bold", False):
                odf_style.addElement(TextProperties(fontweight="bold"))
        if "borders" in style:
            borders = style["borders"]
            for side, thickness in borders.items():
                thickness_translation = {"thin": "0.75pt solid #000000"}
                odf_style.addElement(
                    TableCellProperties(
                        attributes={f"border{side}": thickness_translation[thickness]}
                    )
                )
        if "alignment" in style:
            alignment = style["alignment"]
            horizontal = alignment.get("horizontal")
            if horizontal:
                odf_style.addElement(ParagraphProperties(textalign=horizontal))
            vertical = alignment.get("vertical")
            if vertical:
                odf_style.addElement(TableCellProperties(verticalalign=vertical))
        self.book.styles.addElement(odf_style)
        return name

    def _create_freeze_panes(
        self, sheet_name: str, freeze_panes: tuple[int, int]
    ) -> None:
        """
        Create freeze panes in the sheet.

        Parameters
        ----------
        sheet_name : str
            Name of the spreadsheet
        freeze_panes : tuple of (int, int)
            Freeze pane location x and y
        """
        from odf.config import (
            ConfigItem,
            ConfigItemMapEntry,
            ConfigItemMapIndexed,
            ConfigItemMapNamed,
            ConfigItemSet,
        )

        config_item_set = ConfigItemSet(name="ooo:view-settings")
        self.book.settings.addElement(config_item_set)

        config_item_map_indexed = ConfigItemMapIndexed(name="Views")
        config_item_set.addElement(config_item_map_indexed)

        config_item_map_entry = ConfigItemMapEntry()
        config_item_map_indexed.addElement(config_item_map_entry)

        config_item_map_named = ConfigItemMapNamed(name="Tables")
        config_item_map_entry.addElement(config_item_map_named)

        config_item_map_entry = ConfigItemMapEntry(name=sheet_name)
        config_item_map_named.addElement(config_item_map_entry)

        config_item_map_entry.addElement(
            ConfigItem(name="HorizontalSplitMode", type="short", text="2")
        )
        config_item_map_entry.addElement(
            ConfigItem(name="VerticalSplitMode", type="short", text="2")
        )
        config_item_map_entry.addElement(
            ConfigItem(
                name="HorizontalSplitPosition", type="int", text=str(freeze_panes[0])
            )
        )
        config_item_map_entry.addElement(
            ConfigItem(
                name="VerticalSplitPosition", type="int", text=str(freeze_panes[1])
            )
        )
        config_item_map_entry.addElement(
            ConfigItem(name="PositionRight", type="int", text=str(freeze_panes[0]))
        )
        config_item_map_entry.addElement(
            ConfigItem(name="PositionBottom", type="int", text=str(freeze_panes[1]))
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\ext\mypy\util.py ===
# ext/mypy/util.py
# Copyright (C) 2021-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

from __future__ import annotations

import re
from typing import Any
from typing import Iterable
from typing import Iterator
from typing import List
from typing import Optional
from typing import overload
from typing import Tuple
from typing import Type as TypingType
from typing import TypeVar
from typing import Union

from mypy import version
from mypy.messages import format_type as _mypy_format_type
from mypy.nodes import CallExpr
from mypy.nodes import ClassDef
from mypy.nodes import CLASSDEF_NO_INFO
from mypy.nodes import Context
from mypy.nodes import Expression
from mypy.nodes import FuncDef
from mypy.nodes import IfStmt
from mypy.nodes import JsonDict
from mypy.nodes import MemberExpr
from mypy.nodes import NameExpr
from mypy.nodes import Statement
from mypy.nodes import SymbolTableNode
from mypy.nodes import TypeAlias
from mypy.nodes import TypeInfo
from mypy.options import Options
from mypy.plugin import ClassDefContext
from mypy.plugin import DynamicClassDefContext
from mypy.plugin import SemanticAnalyzerPluginInterface
from mypy.plugins.common import deserialize_and_fixup_type
from mypy.typeops import map_type_from_supertype
from mypy.types import CallableType
from mypy.types import get_proper_type
from mypy.types import Instance
from mypy.types import NoneType
from mypy.types import Type
from mypy.types import TypeVarType
from mypy.types import UnboundType
from mypy.types import UnionType

_vers = tuple(
    [int(x) for x in version.__version__.split(".") if re.match(r"^\d+$", x)]
)
mypy_14 = _vers >= (1, 4)


_TArgType = TypeVar("_TArgType", bound=Union[CallExpr, NameExpr])


class SQLAlchemyAttribute:
    def __init__(
        self,
        name: str,
        line: int,
        column: int,
        typ: Optional[Type],
        info: TypeInfo,
    ) -> None:
        self.name = name
        self.line = line
        self.column = column
        self.type = typ
        self.info = info

    def serialize(self) -> JsonDict:
        assert self.type
        return {
            "name": self.name,
            "line": self.line,
            "column": self.column,
            "type": serialize_type(self.type),
        }

    def expand_typevar_from_subtype(self, sub_type: TypeInfo) -> None:
        """Expands type vars in the context of a subtype when an attribute is
        inherited from a generic super type.
        """
        if not isinstance(self.type, TypeVarType):
            return

        self.type = map_type_from_supertype(self.type, sub_type, self.info)

    @classmethod
    def deserialize(
        cls,
        info: TypeInfo,
        data: JsonDict,
        api: SemanticAnalyzerPluginInterface,
    ) -> SQLAlchemyAttribute:
        data = data.copy()
        typ = deserialize_and_fixup_type(data.pop("type"), api)
        return cls(typ=typ, info=info, **data)


def name_is_dunder(name: str) -> bool:
    return bool(re.match(r"^__.+?__$", name))


def _set_info_metadata(info: TypeInfo, key: str, data: Any) -> None:
    info.metadata.setdefault("sqlalchemy", {})[key] = data


def _get_info_metadata(info: TypeInfo, key: str) -> Optional[Any]:
    return info.metadata.get("sqlalchemy", {}).get(key, None)


def _get_info_mro_metadata(info: TypeInfo, key: str) -> Optional[Any]:
    if info.mro:
        for base in info.mro:
            metadata = _get_info_metadata(base, key)
            if metadata is not None:
                return metadata
    return None


def establish_as_sqlalchemy(info: TypeInfo) -> None:
    info.metadata.setdefault("sqlalchemy", {})


def set_is_base(info: TypeInfo) -> None:
    _set_info_metadata(info, "is_base", True)


def get_is_base(info: TypeInfo) -> bool:
    is_base = _get_info_metadata(info, "is_base")
    return is_base is True


def has_declarative_base(info: TypeInfo) -> bool:
    is_base = _get_info_mro_metadata(info, "is_base")
    return is_base is True


def set_has_table(info: TypeInfo) -> None:
    _set_info_metadata(info, "has_table", True)


def get_has_table(info: TypeInfo) -> bool:
    is_base = _get_info_metadata(info, "has_table")
    return is_base is True


def get_mapped_attributes(
    info: TypeInfo, api: SemanticAnalyzerPluginInterface
) -> Optional[List[SQLAlchemyAttribute]]:
    mapped_attributes: Optional[List[JsonDict]] = _get_info_metadata(
        info, "mapped_attributes"
    )
    if mapped_attributes is None:
        return None

    attributes: List[SQLAlchemyAttribute] = []

    for data in mapped_attributes:
        attr = SQLAlchemyAttribute.deserialize(info, data, api)
        attr.expand_typevar_from_subtype(info)
        attributes.append(attr)

    return attributes


def format_type(typ_: Type, options: Options) -> str:
    if mypy_14:
        return _mypy_format_type(typ_, options)
    else:
        return _mypy_format_type(typ_)  # type: ignore


def set_mapped_attributes(
    info: TypeInfo, attributes: List[SQLAlchemyAttribute]
) -> None:
    _set_info_metadata(
        info,
        "mapped_attributes",
        [attribute.serialize() for attribute in attributes],
    )


def fail(api: SemanticAnalyzerPluginInterface, msg: str, ctx: Context) -> None:
    msg = "[SQLAlchemy Mypy plugin] %s" % msg
    return api.fail(msg, ctx)


def add_global(
    ctx: Union[ClassDefContext, DynamicClassDefContext],
    module: str,
    symbol_name: str,
    asname: str,
) -> None:
    module_globals = ctx.api.modules[ctx.api.cur_mod_id].names

    if asname not in module_globals:
        lookup_sym: SymbolTableNode = ctx.api.modules[module].names[
            symbol_name
        ]

        module_globals[asname] = lookup_sym


@overload
def get_callexpr_kwarg(
    callexpr: CallExpr, name: str, *, expr_types: None = ...
) -> Optional[Union[CallExpr, NameExpr]]: ...


@overload
def get_callexpr_kwarg(
    callexpr: CallExpr,
    name: str,
    *,
    expr_types: Tuple[TypingType[_TArgType], ...],
) -> Optional[_TArgType]: ...


def get_callexpr_kwarg(
    callexpr: CallExpr,
    name: str,
    *,
    expr_types: Optional[Tuple[TypingType[Any], ...]] = None,
) -> Optional[Any]:
    try:
        arg_idx = callexpr.arg_names.index(name)
    except ValueError:
        return None

    kwarg = callexpr.args[arg_idx]
    if isinstance(
        kwarg, expr_types if expr_types is not None else (NameExpr, CallExpr)
    ):
        return kwarg

    return None


def flatten_typechecking(stmts: Iterable[Statement]) -> Iterator[Statement]:
    for stmt in stmts:
        if (
            isinstance(stmt, IfStmt)
            and isinstance(stmt.expr[0], NameExpr)
            and stmt.expr[0].fullname == "typing.TYPE_CHECKING"
        ):
            yield from stmt.body[0].body
        else:
            yield stmt


def type_for_callee(callee: Expression) -> Optional[Union[Instance, TypeInfo]]:
    if isinstance(callee, (MemberExpr, NameExpr)):
        if isinstance(callee.node, FuncDef):
            if callee.node.type and isinstance(callee.node.type, CallableType):
                ret_type = get_proper_type(callee.node.type.ret_type)

                if isinstance(ret_type, Instance):
                    return ret_type

            return None
        elif isinstance(callee.node, TypeAlias):
            target_type = get_proper_type(callee.node.target)
            if isinstance(target_type, Instance):
                return target_type
        elif isinstance(callee.node, TypeInfo):
            return callee.node
    return None


def unbound_to_instance(
    api: SemanticAnalyzerPluginInterface, typ: Type
) -> Type:
    """Take the UnboundType that we seem to get as the ret_type from a FuncDef
    and convert it into an Instance/TypeInfo kind of structure that seems
    to work as the left-hand type of an AssignmentStatement.

    """

    if not isinstance(typ, UnboundType):
        return typ

    # TODO: figure out a more robust way to check this.  The node is some
    # kind of _SpecialForm, there's a typing.Optional that's _SpecialForm,
    # but I can't figure out how to get them to match up
    if typ.name == "Optional":
        # convert from "Optional?" to the more familiar
        # UnionType[..., NoneType()]
        return unbound_to_instance(
            api,
            UnionType(
                [unbound_to_instance(api, typ_arg) for typ_arg in typ.args]
                + [NoneType()]
            ),
        )

    node = api.lookup_qualified(typ.name, typ)

    if (
        node is not None
        and isinstance(node, SymbolTableNode)
        and isinstance(node.node, TypeInfo)
    ):
        bound_type = node.node

        return Instance(
            bound_type,
            [
                (
                    unbound_to_instance(api, arg)
                    if isinstance(arg, UnboundType)
                    else arg
                )
                for arg in typ.args
            ],
        )
    else:
        return typ


def info_for_cls(
    cls: ClassDef, api: SemanticAnalyzerPluginInterface
) -> Optional[TypeInfo]:
    if cls.info is CLASSDEF_NO_INFO:
        sym = api.lookup_qualified(cls.name, cls)
        if sym is None:
            return None
        assert sym and isinstance(sym.node, TypeInfo)
        return sym.node

    return cls.info


def serialize_type(typ: Type) -> Union[str, JsonDict]:
    try:
        return typ.serialize()
    except Exception:
        pass
    if hasattr(typ, "args"):
        typ.args = tuple(
            (
                a.resolve_string_annotation()
                if hasattr(a, "resolve_string_annotation")
                else a
            )
            for a in typ.args
        )
    elif hasattr(typ, "resolve_string_annotation"):
        typ = typ.resolve_string_annotation()
    return typ.serialize()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\codegen\rewriting.py ===
"""
Classes and functions useful for rewriting expressions for optimized code
generation. Some languages (or standards thereof), e.g. C99, offer specialized
math functions for better performance and/or precision.

Using the ``optimize`` function in this module, together with a collection of
rules (represented as instances of ``Optimization``), one can rewrite the
expressions for this purpose::

    >>> from sympy import Symbol, exp, log
    >>> from sympy.codegen.rewriting import optimize, optims_c99
    >>> x = Symbol('x')
    >>> optimize(3*exp(2*x) - 3, optims_c99)
    3*expm1(2*x)
    >>> optimize(exp(2*x) - 1 - exp(-33), optims_c99)
    expm1(2*x) - exp(-33)
    >>> optimize(log(3*x + 3), optims_c99)
    log1p(x) + log(3)
    >>> optimize(log(2*x + 3), optims_c99)
    log(2*x + 3)

The ``optims_c99`` imported above is tuple containing the following instances
(which may be imported from ``sympy.codegen.rewriting``):

- ``expm1_opt``
- ``log1p_opt``
- ``exp2_opt``
- ``log2_opt``
- ``log2const_opt``


"""
from sympy.core.function import expand_log
from sympy.core.singleton import S
from sympy.core.symbol import Wild
from sympy.functions.elementary.complexes import sign
from sympy.functions.elementary.exponential import (exp, log)
from sympy.functions.elementary.miscellaneous import (Max, Min)
from sympy.functions.elementary.trigonometric import (cos, sin, sinc)
from sympy.assumptions import Q, ask
from sympy.codegen.cfunctions import log1p, log2, exp2, expm1
from sympy.codegen.matrix_nodes import MatrixSolve
from sympy.core.expr import UnevaluatedExpr
from sympy.core.power import Pow
from sympy.codegen.numpy_nodes import logaddexp, logaddexp2
from sympy.codegen.scipy_nodes import cosm1, powm1
from sympy.core.mul import Mul
from sympy.matrices.expressions.matexpr import MatrixSymbol
from sympy.utilities.iterables import sift


class Optimization:
    """ Abstract base class for rewriting optimization.

    Subclasses should implement ``__call__`` taking an expression
    as argument.

    Parameters
    ==========
    cost_function : callable returning number
    priority : number

    """
    def __init__(self, cost_function=None, priority=1):
        self.cost_function = cost_function
        self.priority=priority

    def cheapest(self, *args):
        return min(args, key=self.cost_function)


class ReplaceOptim(Optimization):
    """ Rewriting optimization calling replace on expressions.

    Explanation
    ===========

    The instance can be used as a function on expressions for which
    it will apply the ``replace`` method (see
    :meth:`sympy.core.basic.Basic.replace`).

    Parameters
    ==========

    query :
        First argument passed to replace.
    value :
        Second argument passed to replace.

    Examples
    ========

    >>> from sympy import Symbol
    >>> from sympy.codegen.rewriting import ReplaceOptim
    >>> from sympy.codegen.cfunctions import exp2
    >>> x = Symbol('x')
    >>> exp2_opt = ReplaceOptim(lambda p: p.is_Pow and p.base == 2,
    ...     lambda p: exp2(p.exp))
    >>> exp2_opt(2**x)
    exp2(x)

    """

    def __init__(self, query, value, **kwargs):
        super().__init__(**kwargs)
        self.query = query
        self.value = value

    def __call__(self, expr):
        return expr.replace(self.query, self.value)


def optimize(expr, optimizations):
    """ Apply optimizations to an expression.

    Parameters
    ==========

    expr : expression
    optimizations : iterable of ``Optimization`` instances
        The optimizations will be sorted with respect to ``priority`` (highest first).

    Examples
    ========

    >>> from sympy import log, Symbol
    >>> from sympy.codegen.rewriting import optims_c99, optimize
    >>> x = Symbol('x')
    >>> optimize(log(x+3)/log(2) + log(x**2 + 1), optims_c99)
    log1p(x**2) + log2(x + 3)

    """

    for optim in sorted(optimizations, key=lambda opt: opt.priority, reverse=True):
        new_expr = optim(expr)
        if optim.cost_function is None:
            expr = new_expr
        else:
            expr = optim.cheapest(expr, new_expr)
    return expr


exp2_opt = ReplaceOptim(
    lambda p: p.is_Pow and p.base == 2,
    lambda p: exp2(p.exp)
)


_d = Wild('d', properties=[lambda x: x.is_Dummy])
_u = Wild('u', properties=[lambda x: not x.is_number and not x.is_Add])
_v = Wild('v')
_w = Wild('w')
_n = Wild('n', properties=[lambda x: x.is_number])

sinc_opt1 = ReplaceOptim(
    sin(_w)/_w, sinc(_w)
)
sinc_opt2 = ReplaceOptim(
    sin(_n*_w)/_w, _n*sinc(_n*_w)
)
sinc_opts = (sinc_opt1, sinc_opt2)

log2_opt = ReplaceOptim(_v*log(_w)/log(2), _v*log2(_w), cost_function=lambda expr: expr.count(
    lambda e: (  # division & eval of transcendentals are expensive floating point operations...
        e.is_Pow and e.exp.is_negative  # division
        or (isinstance(e, (log, log2)) and not e.args[0].is_number))  # transcendental
    )
)

log2const_opt = ReplaceOptim(log(2)*log2(_w), log(_w))

logsumexp_2terms_opt = ReplaceOptim(
    lambda l: (isinstance(l, log)
               and l.args[0].is_Add
               and len(l.args[0].args) == 2
               and all(isinstance(t, exp) for t in l.args[0].args)),
    lambda l: (
        Max(*[e.args[0] for e in l.args[0].args]) +
        log1p(exp(Min(*[e.args[0] for e in l.args[0].args])))
    )
)


class FuncMinusOneOptim(ReplaceOptim):
    """Specialization of ReplaceOptim for functions evaluating "f(x) - 1".

    Explanation
    ===========

    Numerical functions which go toward one as x go toward zero is often best
    implemented by a dedicated function in order to avoid catastrophic
    cancellation. One such example is ``expm1(x)`` in the C standard library
    which evaluates ``exp(x) - 1``. Such functions preserves many more
    significant digits when its argument is much smaller than one, compared
    to subtracting one afterwards.

    Parameters
    ==========

    func :
        The function which is subtracted by one.
    func_m_1 :
        The specialized function evaluating ``func(x) - 1``.
    opportunistic : bool
        When ``True``, apply the transformation as long as the magnitude of the
        remaining number terms decreases. When ``False``, only apply the
        transformation if it completely eliminates the number term.

    Examples
    ========

    >>> from sympy import symbols, exp
    >>> from sympy.codegen.rewriting import FuncMinusOneOptim
    >>> from sympy.codegen.cfunctions import expm1
    >>> x, y = symbols('x y')
    >>> expm1_opt = FuncMinusOneOptim(exp, expm1)
    >>> expm1_opt(exp(x) + 2*exp(5*y) - 3)
    expm1(x) + 2*expm1(5*y)


    """

    def __init__(self, func, func_m_1, opportunistic=True):
        weight = 10  # <-- this is an arbitrary number (heuristic)
        super().__init__(lambda e: e.is_Add, self.replace_in_Add,
                         cost_function=lambda expr: expr.count_ops() - weight*expr.count(func_m_1))
        self.func = func
        self.func_m_1 = func_m_1
        self.opportunistic = opportunistic

    def _group_Add_terms(self, add):
        numbers, non_num = sift(add.args, lambda arg: arg.is_number, binary=True)
        numsum = sum(numbers)
        terms_with_func, other = sift(non_num, lambda arg: arg.has(self.func), binary=True)
        return numsum, terms_with_func, other

    def replace_in_Add(self, e):
        """ passed as second argument to Basic.replace(...) """
        numsum, terms_with_func, other_non_num_terms = self._group_Add_terms(e)
        if numsum == 0:
            return e
        substituted, untouched = [], []
        for with_func in terms_with_func:
            if with_func.is_Mul:
                func, coeff = sift(with_func.args, lambda arg: arg.func == self.func, binary=True)
                if len(func) == 1 and len(coeff) == 1:
                    func, coeff = func[0], coeff[0]
                else:
                    coeff = None
            elif with_func.func == self.func:
                func, coeff = with_func, S.One
            else:
                coeff = None

            if coeff is not None and coeff.is_number and sign(coeff) == -sign(numsum):
                if self.opportunistic:
                    do_substitute = abs(coeff+numsum) < abs(numsum)
                else:
                    do_substitute = coeff+numsum == 0

                if do_substitute:  # advantageous substitution
                    numsum += coeff
                    substituted.append(coeff*self.func_m_1(*func.args))
                    continue
            untouched.append(with_func)

        return e.func(numsum, *substituted, *untouched, *other_non_num_terms)

    def __call__(self, expr):
        alt1 = super().__call__(expr)
        alt2 = super().__call__(expr.factor())
        return self.cheapest(alt1, alt2)


expm1_opt = FuncMinusOneOptim(exp, expm1)
cosm1_opt = FuncMinusOneOptim(cos, cosm1)
powm1_opt = FuncMinusOneOptim(Pow, powm1)

log1p_opt = ReplaceOptim(
    lambda e: isinstance(e, log),
    lambda l: expand_log(l.replace(
        log, lambda arg: log(arg.factor())
    )).replace(log(_u+1), log1p(_u))
)

def create_expand_pow_optimization(limit, *, base_req=lambda b: b.is_symbol):
    """ Creates an instance of :class:`ReplaceOptim` for expanding ``Pow``.

    Explanation
    ===========

    The requirements for expansions are that the base needs to be a symbol
    and the exponent needs to be an Integer (and be less than or equal to
    ``limit``).

    Parameters
    ==========

    limit : int
         The highest power which is expanded into multiplication.
    base_req : function returning bool
         Requirement on base for expansion to happen, default is to return
         the ``is_symbol`` attribute of the base.

    Examples
    ========

    >>> from sympy import Symbol, sin
    >>> from sympy.codegen.rewriting import create_expand_pow_optimization
    >>> x = Symbol('x')
    >>> expand_opt = create_expand_pow_optimization(3)
    >>> expand_opt(x**5 + x**3)
    x**5 + x*x*x
    >>> expand_opt(x**5 + x**3 + sin(x)**3)
    x**5 + sin(x)**3 + x*x*x
    >>> opt2 = create_expand_pow_optimization(3, base_req=lambda b: not b.is_Function)
    >>> opt2((x+1)**2 + sin(x)**2)
    sin(x)**2 + (x + 1)*(x + 1)

    """
    return ReplaceOptim(
        lambda e: e.is_Pow and base_req(e.base) and e.exp.is_Integer and abs(e.exp) <= limit,
        lambda p: (
            UnevaluatedExpr(Mul(*([p.base]*+p.exp), evaluate=False)) if p.exp > 0 else
            1/UnevaluatedExpr(Mul(*([p.base]*-p.exp), evaluate=False))
        ))

# Optimization procedures for turning A**(-1) * x into MatrixSolve(A, x)
def _matinv_predicate(expr):
    # TODO: We should be able to support more than 2 elements
    if expr.is_MatMul and len(expr.args) == 2:
        left, right = expr.args
        if left.is_Inverse and right.shape[1] == 1:
            inv_arg = left.arg
            if isinstance(inv_arg, MatrixSymbol):
                return bool(ask(Q.fullrank(left.arg)))

    return False

def _matinv_transform(expr):
    left, right = expr.args
    inv_arg = left.arg
    return MatrixSolve(inv_arg, right)


matinv_opt = ReplaceOptim(_matinv_predicate, _matinv_transform)


logaddexp_opt = ReplaceOptim(log(exp(_v)+exp(_w)), logaddexp(_v, _w))
logaddexp2_opt = ReplaceOptim(log(Pow(2, _v)+Pow(2, _w)), logaddexp2(_v, _w)*log(2))

# Collections of optimizations:
optims_c99 = (expm1_opt, log1p_opt, exp2_opt, log2_opt, log2const_opt)

optims_numpy = optims_c99 + (logaddexp_opt, logaddexp2_opt,) + sinc_opts

optims_scipy = (cosm1_opt, powm1_opt)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\combinatorics\testutil.py ===
from sympy.combinatorics import Permutation
from sympy.combinatorics.util import _distribute_gens_by_base

rmul = Permutation.rmul


def _cmp_perm_lists(first, second):
    """
    Compare two lists of permutations as sets.

    Explanation
    ===========

    This is used for testing purposes. Since the array form of a
    permutation is currently a list, Permutation is not hashable
    and cannot be put into a set.

    Examples
    ========

    >>> from sympy.combinatorics.permutations import Permutation
    >>> from sympy.combinatorics.testutil import _cmp_perm_lists
    >>> a = Permutation([0, 2, 3, 4, 1])
    >>> b = Permutation([1, 2, 0, 4, 3])
    >>> c = Permutation([3, 4, 0, 1, 2])
    >>> ls1 = [a, b, c]
    >>> ls2 = [b, c, a]
    >>> _cmp_perm_lists(ls1, ls2)
    True

    """
    return {tuple(a) for a in first} == \
           {tuple(a) for a in second}


def _naive_list_centralizer(self, other, af=False):
    from sympy.combinatorics.perm_groups import PermutationGroup
    """
    Return a list of elements for the centralizer of a subgroup/set/element.

    Explanation
    ===========

    This is a brute force implementation that goes over all elements of the
    group and checks for membership in the centralizer. It is used to
    test ``.centralizer()`` from ``sympy.combinatorics.perm_groups``.

    Examples
    ========

    >>> from sympy.combinatorics.testutil import _naive_list_centralizer
    >>> from sympy.combinatorics.named_groups import DihedralGroup
    >>> D = DihedralGroup(4)
    >>> _naive_list_centralizer(D, D)
    [Permutation([0, 1, 2, 3]), Permutation([2, 3, 0, 1])]

    See Also
    ========

    sympy.combinatorics.perm_groups.centralizer

    """
    from sympy.combinatorics.permutations import _af_commutes_with
    if hasattr(other, 'generators'):
        elements = list(self.generate_dimino(af=True))
        gens = [x._array_form for x in other.generators]
        commutes_with_gens = lambda x: all(_af_commutes_with(x, gen) for gen in gens)
        centralizer_list = []
        if not af:
            for element in elements:
                if commutes_with_gens(element):
                    centralizer_list.append(Permutation._af_new(element))
        else:
            for element in elements:
                if commutes_with_gens(element):
                    centralizer_list.append(element)
        return centralizer_list
    elif hasattr(other, 'getitem'):
        return _naive_list_centralizer(self, PermutationGroup(other), af)
    elif hasattr(other, 'array_form'):
        return _naive_list_centralizer(self, PermutationGroup([other]), af)


def _verify_bsgs(group, base, gens):
    """
    Verify the correctness of a base and strong generating set.

    Explanation
    ===========

    This is a naive implementation using the definition of a base and a strong
    generating set relative to it. There are other procedures for
    verifying a base and strong generating set, but this one will
    serve for more robust testing.

    Examples
    ========

    >>> from sympy.combinatorics.named_groups import AlternatingGroup
    >>> from sympy.combinatorics.testutil import _verify_bsgs
    >>> A = AlternatingGroup(4)
    >>> A.schreier_sims()
    >>> _verify_bsgs(A, A.base, A.strong_gens)
    True

    See Also
    ========

    sympy.combinatorics.perm_groups.PermutationGroup.schreier_sims

    """
    from sympy.combinatorics.perm_groups import PermutationGroup
    strong_gens_distr = _distribute_gens_by_base(base, gens)
    current_stabilizer = group
    for i in range(len(base)):
        candidate = PermutationGroup(strong_gens_distr[i])
        if current_stabilizer.order() != candidate.order():
            return False
        current_stabilizer = current_stabilizer.stabilizer(base[i])
    if current_stabilizer.order() != 1:
        return False
    return True


def _verify_centralizer(group, arg, centr=None):
    """
    Verify the centralizer of a group/set/element inside another group.

    This is used for testing ``.centralizer()`` from
    ``sympy.combinatorics.perm_groups``

    Examples
    ========

    >>> from sympy.combinatorics.named_groups import (SymmetricGroup,
    ... AlternatingGroup)
    >>> from sympy.combinatorics.perm_groups import PermutationGroup
    >>> from sympy.combinatorics.permutations import Permutation
    >>> from sympy.combinatorics.testutil import _verify_centralizer
    >>> S = SymmetricGroup(5)
    >>> A = AlternatingGroup(5)
    >>> centr = PermutationGroup([Permutation([0, 1, 2, 3, 4])])
    >>> _verify_centralizer(S, A, centr)
    True

    See Also
    ========

    _naive_list_centralizer,
    sympy.combinatorics.perm_groups.PermutationGroup.centralizer,
    _cmp_perm_lists

    """
    if centr is None:
        centr = group.centralizer(arg)
    centr_list = list(centr.generate_dimino(af=True))
    centr_list_naive = _naive_list_centralizer(group, arg, af=True)
    return _cmp_perm_lists(centr_list, centr_list_naive)


def _verify_normal_closure(group, arg, closure=None):
    from sympy.combinatorics.perm_groups import PermutationGroup
    """
    Verify the normal closure of a subgroup/subset/element in a group.

    This is used to test
    sympy.combinatorics.perm_groups.PermutationGroup.normal_closure

    Examples
    ========

    >>> from sympy.combinatorics.named_groups import (SymmetricGroup,
    ... AlternatingGroup)
    >>> from sympy.combinatorics.testutil import _verify_normal_closure
    >>> S = SymmetricGroup(3)
    >>> A = AlternatingGroup(3)
    >>> _verify_normal_closure(S, A, closure=A)
    True

    See Also
    ========

    sympy.combinatorics.perm_groups.PermutationGroup.normal_closure

    """
    if closure is None:
        closure = group.normal_closure(arg)
    conjugates = set()
    if hasattr(arg, 'generators'):
        subgr_gens = arg.generators
    elif hasattr(arg, '__getitem__'):
        subgr_gens = arg
    elif hasattr(arg, 'array_form'):
        subgr_gens = [arg]
    for el in group.generate_dimino():
        conjugates.update(gen ^ el for gen in subgr_gens)
    naive_closure = PermutationGroup(list(conjugates))
    return closure.is_subgroup(naive_closure)


def canonicalize_naive(g, dummies, sym, *v):
    """
    Canonicalize tensor formed by tensors of the different types.

    Explanation
    ===========

    sym_i symmetry under exchange of two component tensors of type `i`
          None  no symmetry
          0     commuting
          1     anticommuting

    Parameters
    ==========

    g : Permutation representing the tensor.
    dummies : List of dummy indices.
    msym : Symmetry of the metric.
    v : A list of (base_i, gens_i, n_i, sym_i) for tensors of type `i`.
        base_i, gens_i BSGS for tensors of this type
        n_i  number of tensors of type `i`

    Returns
    =======

    Returns 0 if the tensor is zero, else returns the array form of
    the permutation representing the canonical form of the tensor.

    Examples
    ========

    >>> from sympy.combinatorics.testutil import canonicalize_naive
    >>> from sympy.combinatorics.tensor_can import get_symmetric_group_sgs
    >>> from sympy.combinatorics import Permutation
    >>> g = Permutation([1, 3, 2, 0, 4, 5])
    >>> base2, gens2 = get_symmetric_group_sgs(2)
    >>> canonicalize_naive(g, [2, 3], 0, (base2, gens2, 2, 0))
    [0, 2, 1, 3, 4, 5]
    """
    from sympy.combinatorics.perm_groups import PermutationGroup
    from sympy.combinatorics.tensor_can import gens_products, dummy_sgs
    from sympy.combinatorics.permutations import _af_rmul
    v1 = []
    for i in range(len(v)):
        base_i, gens_i, n_i, sym_i = v[i]
        v1.append((base_i, gens_i, [[]]*n_i, sym_i))
    size, sbase, sgens = gens_products(*v1)
    dgens = dummy_sgs(dummies, sym, size-2)
    if isinstance(sym, int):
        num_types = 1
        dummies = [dummies]
        sym = [sym]
    else:
        num_types = len(sym)
    dgens = []
    for i in range(num_types):
        dgens.extend(dummy_sgs(dummies[i], sym[i], size - 2))
    S = PermutationGroup(sgens)
    D = PermutationGroup([Permutation(x) for x in dgens])
    dlist = list(D.generate(af=True))
    g = g.array_form
    st = set()
    for s in S.generate(af=True):
        h = _af_rmul(g, s)
        for d in dlist:
            q = tuple(_af_rmul(d, h))
            st.add(q)
    a = list(st)
    a.sort()
    prev = (0,)*size
    for h in a:
        if h[:-2] == prev[:-2]:
            if h[-1] != prev[-1]:
                return 0
        prev = h
    return list(a[0])


def graph_certificate(gr):
    """
    Return a certificate for the graph

    Parameters
    ==========

    gr : adjacency list

    Explanation
    ===========

    The graph is assumed to be unoriented and without
    external lines.

    Associate to each vertex of the graph a symmetric tensor with
    number of indices equal to the degree of the vertex; indices
    are contracted when they correspond to the same line of the graph.
    The canonical form of the tensor gives a certificate for the graph.

    This is not an efficient algorithm to get the certificate of a graph.

    Examples
    ========

    >>> from sympy.combinatorics.testutil import graph_certificate
    >>> gr1 = {0:[1, 2, 3, 5], 1:[0, 2, 4], 2:[0, 1, 3, 4], 3:[0, 2, 4], 4:[1, 2, 3, 5], 5:[0, 4]}
    >>> gr2 = {0:[1, 5], 1:[0, 2, 3, 4], 2:[1, 3, 5], 3:[1, 2, 4, 5], 4:[1, 3, 5], 5:[0, 2, 3, 4]}
    >>> c1 = graph_certificate(gr1)
    >>> c2 = graph_certificate(gr2)
    >>> c1
    [0, 2, 4, 6, 1, 8, 10, 12, 3, 14, 16, 18, 5, 9, 15, 7, 11, 17, 13, 19, 20, 21]
    >>> c1 == c2
    True
    """
    from sympy.combinatorics.permutations import _af_invert
    from sympy.combinatorics.tensor_can import get_symmetric_group_sgs, canonicalize
    items = list(gr.items())
    items.sort(key=lambda x: len(x[1]), reverse=True)
    pvert = [x[0] for x in items]
    pvert = _af_invert(pvert)

    # the indices of the tensor are twice the number of lines of the graph
    num_indices = 0
    for v, neigh in items:
        num_indices += len(neigh)
    # associate to each vertex its indices; for each line
    # between two vertices assign the
    # even index to the vertex which comes first in items,
    # the odd index to the other vertex
    vertices = [[] for i in items]
    i = 0
    for v, neigh in items:
        for v2 in neigh:
            if pvert[v] < pvert[v2]:
                vertices[pvert[v]].append(i)
                vertices[pvert[v2]].append(i+1)
                i += 2
    g = []
    for v in vertices:
        g.extend(v)
    assert len(g) == num_indices
    g += [num_indices, num_indices + 1]
    size = num_indices + 2
    assert sorted(g) == list(range(size))
    g = Permutation(g)
    vlen = [0]*(len(vertices[0])+1)
    for neigh in vertices:
        vlen[len(neigh)] += 1
    v = []
    for i in range(len(vlen)):
        n = vlen[i]
        if n:
            base, gens = get_symmetric_group_sgs(i)
            v.append((base, gens, n, 0))
    v.reverse()
    dummies = list(range(num_indices))
    can = canonicalize(g, dummies, 0, *v)
    return can

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\arrays\_arrow_string_mixins.py ===
from __future__ import annotations

from functools import partial
import re
from typing import (
    TYPE_CHECKING,
    Any,
    Literal,
)

import numpy as np

from pandas._libs import lib
from pandas.compat import (
    pa_version_under10p1,
    pa_version_under11p0,
    pa_version_under13p0,
    pa_version_under17p0,
)

if not pa_version_under10p1:
    import pyarrow as pa
    import pyarrow.compute as pc

if TYPE_CHECKING:
    from collections.abc import Callable

    from pandas._typing import (
        Scalar,
        Self,
    )


class ArrowStringArrayMixin:
    _pa_array: pa.ChunkedArray

    def __init__(self, *args, **kwargs) -> None:
        raise NotImplementedError

    def _convert_bool_result(self, result, na=lib.no_default, method_name=None):
        # Convert a bool-dtype result to the appropriate result type
        raise NotImplementedError

    def _convert_int_result(self, result):
        # Convert an integer-dtype result to the appropriate result type
        raise NotImplementedError

    def _apply_elementwise(self, func: Callable) -> list[list[Any]]:
        raise NotImplementedError

    def _str_len(self):
        result = pc.utf8_length(self._pa_array)
        return self._convert_int_result(result)

    def _str_lower(self) -> Self:
        return type(self)(pc.utf8_lower(self._pa_array))

    def _str_upper(self) -> Self:
        return type(self)(pc.utf8_upper(self._pa_array))

    def _str_strip(self, to_strip=None) -> Self:
        if to_strip is None:
            result = pc.utf8_trim_whitespace(self._pa_array)
        else:
            result = pc.utf8_trim(self._pa_array, characters=to_strip)
        return type(self)(result)

    def _str_lstrip(self, to_strip=None) -> Self:
        if to_strip is None:
            result = pc.utf8_ltrim_whitespace(self._pa_array)
        else:
            result = pc.utf8_ltrim(self._pa_array, characters=to_strip)
        return type(self)(result)

    def _str_rstrip(self, to_strip=None) -> Self:
        if to_strip is None:
            result = pc.utf8_rtrim_whitespace(self._pa_array)
        else:
            result = pc.utf8_rtrim(self._pa_array, characters=to_strip)
        return type(self)(result)

    def _str_pad(
        self,
        width: int,
        side: Literal["left", "right", "both"] = "left",
        fillchar: str = " ",
    ):
        if side == "left":
            pa_pad = pc.utf8_lpad
        elif side == "right":
            pa_pad = pc.utf8_rpad
        elif side == "both":
            if pa_version_under17p0:
                # GH#59624 fall back to object dtype
                from pandas import array as pd_array

                obj_arr = self.astype(object, copy=False)  # type: ignore[attr-defined]
                obj = pd_array(obj_arr, dtype=object)
                result = obj._str_pad(width, side, fillchar)  # type: ignore[attr-defined]
                return type(self)._from_sequence(result, dtype=self.dtype)  # type: ignore[attr-defined]
            else:
                # GH#54792
                # https://github.com/apache/arrow/issues/15053#issuecomment-2317032347
                lean_left = (width % 2) == 0
                pa_pad = partial(pc.utf8_center, lean_left_on_odd_padding=lean_left)
        else:
            raise ValueError(
                f"Invalid side: {side}. Side must be one of 'left', 'right', 'both'"
            )
        return type(self)(pa_pad(self._pa_array, width=width, padding=fillchar))

    def _str_get(self, i: int):
        lengths = pc.utf8_length(self._pa_array)
        if i >= 0:
            out_of_bounds = pc.greater_equal(i, lengths)
            start = i
            stop = i + 1
            step = 1
        else:
            out_of_bounds = pc.greater(-i, lengths)
            start = i
            stop = i - 1
            step = -1
        not_out_of_bounds = pc.invert(out_of_bounds.fill_null(True))
        selected = pc.utf8_slice_codeunits(
            self._pa_array, start=start, stop=stop, step=step
        )
        null_value = pa.scalar(None, type=self._pa_array.type)
        result = pc.if_else(not_out_of_bounds, selected, null_value)
        return type(self)(result)

    def _str_slice(
        self, start: int | None = None, stop: int | None = None, step: int | None = None
    ):
        if pa_version_under11p0:
            # GH#59724
            result = self._apply_elementwise(lambda val: val[start:stop:step])
            return type(self)(pa.chunked_array(result, type=self._pa_array.type))
        if start is None:
            if step is not None and step < 0:
                # GH#59710
                start = -1
            else:
                start = 0
        if step is None:
            step = 1
        return type(self)(
            pc.utf8_slice_codeunits(self._pa_array, start=start, stop=stop, step=step)
        )

    def _str_slice_replace(
        self, start: int | None = None, stop: int | None = None, repl: str | None = None
    ):
        if repl is None:
            repl = ""
        if start is None:
            start = 0
        if stop is None:
            stop = np.iinfo(np.int64).max
        return type(self)(pc.utf8_replace_slice(self._pa_array, start, stop, repl))

    def _str_replace(
        self,
        pat: str | re.Pattern,
        repl: str | Callable,
        n: int = -1,
        case: bool = True,
        flags: int = 0,
        regex: bool = True,
    ) -> Self:
        if isinstance(pat, re.Pattern) or callable(repl) or not case or flags:
            raise NotImplementedError(
                "replace is not supported with a re.Pattern, callable repl, "
                "case=False, or flags!=0"
            )

        func = pc.replace_substring_regex if regex else pc.replace_substring
        # https://github.com/apache/arrow/issues/39149
        # GH 56404, unexpected behavior with negative max_replacements with pyarrow.
        pa_max_replacements = None if n < 0 else n
        result = func(
            self._pa_array,
            pattern=pat,
            replacement=repl,
            max_replacements=pa_max_replacements,
        )
        return type(self)(result)

    def _str_capitalize(self) -> Self:
        return type(self)(pc.utf8_capitalize(self._pa_array))

    def _str_title(self):
        return type(self)(pc.utf8_title(self._pa_array))

    def _str_swapcase(self):
        return type(self)(pc.utf8_swapcase(self._pa_array))

    def _str_removeprefix(self, prefix: str):
        if not pa_version_under13p0:
            starts_with = pc.starts_with(self._pa_array, pattern=prefix)
            removed = pc.utf8_slice_codeunits(self._pa_array, len(prefix))
            result = pc.if_else(starts_with, removed, self._pa_array)
            return type(self)(result)
        predicate = lambda val: val.removeprefix(prefix)
        result = self._apply_elementwise(predicate)
        return type(self)(pa.chunked_array(result))

    def _str_removesuffix(self, suffix: str):
        ends_with = pc.ends_with(self._pa_array, pattern=suffix)
        removed = pc.utf8_slice_codeunits(self._pa_array, 0, stop=-len(suffix))
        result = pc.if_else(ends_with, removed, self._pa_array)
        return type(self)(result)

    def _str_startswith(
        self, pat: str | tuple[str, ...], na: Scalar | lib.NoDefault = lib.no_default
    ):
        if isinstance(pat, str):
            result = pc.starts_with(self._pa_array, pattern=pat)
        else:
            if len(pat) == 0:
                # For empty tuple we return null for missing values and False
                #  for valid values.
                result = pc.if_else(pc.is_null(self._pa_array), None, False)
            else:
                result = pc.starts_with(self._pa_array, pattern=pat[0])

                for p in pat[1:]:
                    result = pc.or_(result, pc.starts_with(self._pa_array, pattern=p))
        return self._convert_bool_result(result, na=na, method_name="startswith")

    def _str_endswith(
        self, pat: str | tuple[str, ...], na: Scalar | lib.NoDefault = lib.no_default
    ):
        if isinstance(pat, str):
            result = pc.ends_with(self._pa_array, pattern=pat)
        else:
            if len(pat) == 0:
                # For empty tuple we return null for missing values and False
                #  for valid values.
                result = pc.if_else(pc.is_null(self._pa_array), None, False)
            else:
                result = pc.ends_with(self._pa_array, pattern=pat[0])

                for p in pat[1:]:
                    result = pc.or_(result, pc.ends_with(self._pa_array, pattern=p))
        return self._convert_bool_result(result, na=na, method_name="endswith")

    def _str_isalnum(self):
        result = pc.utf8_is_alnum(self._pa_array)
        return self._convert_bool_result(result)

    def _str_isalpha(self):
        result = pc.utf8_is_alpha(self._pa_array)
        return self._convert_bool_result(result)

    def _str_isdecimal(self):
        result = pc.utf8_is_decimal(self._pa_array)
        return self._convert_bool_result(result)

    def _str_isdigit(self):
        result = pc.utf8_is_digit(self._pa_array)
        return self._convert_bool_result(result)

    def _str_islower(self):
        result = pc.utf8_is_lower(self._pa_array)
        return self._convert_bool_result(result)

    def _str_isnumeric(self):
        result = pc.utf8_is_numeric(self._pa_array)
        return self._convert_bool_result(result)

    def _str_isspace(self):
        result = pc.utf8_is_space(self._pa_array)
        return self._convert_bool_result(result)

    def _str_istitle(self):
        result = pc.utf8_is_title(self._pa_array)
        return self._convert_bool_result(result)

    def _str_isupper(self):
        result = pc.utf8_is_upper(self._pa_array)
        return self._convert_bool_result(result)

    def _str_contains(
        self,
        pat,
        case: bool = True,
        flags: int = 0,
        na: Scalar | lib.NoDefault = lib.no_default,
        regex: bool = True,
    ):
        if flags:
            raise NotImplementedError(f"contains not implemented with {flags=}")

        if regex:
            pa_contains = pc.match_substring_regex
        else:
            pa_contains = pc.match_substring
        result = pa_contains(self._pa_array, pat, ignore_case=not case)
        return self._convert_bool_result(result, na=na, method_name="contains")

    def _str_match(
        self,
        pat: str,
        case: bool = True,
        flags: int = 0,
        na: Scalar | lib.NoDefault = lib.no_default,
    ):
        if not pat.startswith("^"):
            pat = f"^{pat}"
        return self._str_contains(pat, case, flags, na, regex=True)

    def _str_fullmatch(
        self,
        pat,
        case: bool = True,
        flags: int = 0,
        na: Scalar | lib.NoDefault = lib.no_default,
    ):
        if not pat.endswith("$") or pat.endswith("\\$"):
            pat = f"{pat}$"
        return self._str_match(pat, case, flags, na)

    def _str_find(self, sub: str, start: int = 0, end: int | None = None):
        if (
            pa_version_under13p0
            and not (start != 0 and end is not None)
            and not (start == 0 and end is None)
        ):
            # GH#59562
            res_list = self._apply_elementwise(lambda val: val.find(sub, start, end))
            return self._convert_int_result(pa.chunked_array(res_list))

        if (start == 0 or start is None) and end is None:
            result = pc.find_substring(self._pa_array, sub)
        else:
            if sub == "":
                # GH#56792
                res_list = self._apply_elementwise(
                    lambda val: val.find(sub, start, end)
                )
                return self._convert_int_result(pa.chunked_array(res_list))
            if start is None:
                start_offset = 0
                start = 0
            elif start < 0:
                start_offset = pc.add(start, pc.utf8_length(self._pa_array))
                start_offset = pc.if_else(pc.less(start_offset, 0), 0, start_offset)
            else:
                start_offset = start
            slices = pc.utf8_slice_codeunits(self._pa_array, start, stop=end)
            result = pc.find_substring(slices, sub)
            found = pc.not_equal(result, pa.scalar(-1, type=result.type))
            offset_result = pc.add(result, start_offset)
            result = pc.if_else(found, offset_result, -1)
        return self._convert_int_result(result)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\indexes\timedeltas.py ===
""" implement the TimedeltaIndex """
from __future__ import annotations

from typing import TYPE_CHECKING
import warnings

from pandas._libs import (
    index as libindex,
    lib,
)
from pandas._libs.tslibs import (
    Resolution,
    Timedelta,
    to_offset,
)
from pandas._libs.tslibs.timedeltas import disallow_ambiguous_unit
from pandas.util._exceptions import find_stack_level

from pandas.core.dtypes.common import (
    is_scalar,
    pandas_dtype,
)
from pandas.core.dtypes.generic import ABCSeries

from pandas.core.arrays.timedeltas import TimedeltaArray
import pandas.core.common as com
from pandas.core.indexes.base import (
    Index,
    maybe_extract_name,
)
from pandas.core.indexes.datetimelike import DatetimeTimedeltaMixin
from pandas.core.indexes.extension import inherit_names

if TYPE_CHECKING:
    from pandas._typing import DtypeObj


@inherit_names(
    ["__neg__", "__pos__", "__abs__", "total_seconds", "round", "floor", "ceil"]
    + TimedeltaArray._field_ops,
    TimedeltaArray,
    wrap=True,
)
@inherit_names(
    [
        "components",
        "to_pytimedelta",
        "sum",
        "std",
        "median",
    ],
    TimedeltaArray,
)
class TimedeltaIndex(DatetimeTimedeltaMixin):
    """
    Immutable Index of timedelta64 data.

    Represented internally as int64, and scalars returned Timedelta objects.

    Parameters
    ----------
    data : array-like (1-dimensional), optional
        Optional timedelta-like data to construct index with.
    unit : {'D', 'h', 'm', 's', 'ms', 'us', 'ns'}, optional
        The unit of ``data``.

        .. deprecated:: 2.2.0
         Use ``pd.to_timedelta`` instead.

    freq : str or pandas offset object, optional
        One of pandas date offset strings or corresponding objects. The string
        ``'infer'`` can be passed in order to set the frequency of the index as
        the inferred frequency upon creation.
    dtype : numpy.dtype or str, default None
        Valid ``numpy`` dtypes are ``timedelta64[ns]``, ``timedelta64[us]``,
        ``timedelta64[ms]``, and ``timedelta64[s]``.
    copy : bool
        Make a copy of input array.
    name : object
        Name to be stored in the index.

    Attributes
    ----------
    days
    seconds
    microseconds
    nanoseconds
    components
    inferred_freq

    Methods
    -------
    to_pytimedelta
    to_series
    round
    floor
    ceil
    to_frame
    mean

    See Also
    --------
    Index : The base pandas Index type.
    Timedelta : Represents a duration between two dates or times.
    DatetimeIndex : Index of datetime64 data.
    PeriodIndex : Index of Period data.
    timedelta_range : Create a fixed-frequency TimedeltaIndex.

    Notes
    -----
    To learn more about the frequency strings, please see `this link
    <https://pandas.pydata.org/pandas-docs/stable/user_guide/timeseries.html#offset-aliases>`__.

    Examples
    --------
    >>> pd.TimedeltaIndex(['0 days', '1 days', '2 days', '3 days', '4 days'])
    TimedeltaIndex(['0 days', '1 days', '2 days', '3 days', '4 days'],
                   dtype='timedelta64[ns]', freq=None)

    We can also let pandas infer the frequency when possible.

    >>> pd.TimedeltaIndex(np.arange(5) * 24 * 3600 * 1e9, freq='infer')
    TimedeltaIndex(['0 days', '1 days', '2 days', '3 days', '4 days'],
                   dtype='timedelta64[ns]', freq='D')
    """

    _typ = "timedeltaindex"

    _data_cls = TimedeltaArray

    @property
    def _engine_type(self) -> type[libindex.TimedeltaEngine]:
        return libindex.TimedeltaEngine

    _data: TimedeltaArray

    # Use base class method instead of DatetimeTimedeltaMixin._get_string_slice
    _get_string_slice = Index._get_string_slice

    # error: Signature of "_resolution_obj" incompatible with supertype
    # "DatetimeIndexOpsMixin"
    @property
    def _resolution_obj(self) -> Resolution | None:  # type: ignore[override]
        return self._data._resolution_obj

    # -------------------------------------------------------------------
    # Constructors

    def __new__(
        cls,
        data=None,
        unit=lib.no_default,
        freq=lib.no_default,
        closed=lib.no_default,
        dtype=None,
        copy: bool = False,
        name=None,
    ):
        if closed is not lib.no_default:
            # GH#52628
            warnings.warn(
                f"The 'closed' keyword in {cls.__name__} construction is "
                "deprecated and will be removed in a future version.",
                FutureWarning,
                stacklevel=find_stack_level(),
            )

        if unit is not lib.no_default:
            # GH#55499
            warnings.warn(
                f"The 'unit' keyword in {cls.__name__} construction is "
                "deprecated and will be removed in a future version. "
                "Use pd.to_timedelta instead.",
                FutureWarning,
                stacklevel=find_stack_level(),
            )
        else:
            unit = None

        name = maybe_extract_name(name, data, cls)

        if is_scalar(data):
            cls._raise_scalar_data_error(data)

        disallow_ambiguous_unit(unit)
        if dtype is not None:
            dtype = pandas_dtype(dtype)

        if (
            isinstance(data, TimedeltaArray)
            and freq is lib.no_default
            and (dtype is None or dtype == data.dtype)
        ):
            if copy:
                data = data.copy()
            return cls._simple_new(data, name=name)

        if (
            isinstance(data, TimedeltaIndex)
            and freq is lib.no_default
            and name is None
            and (dtype is None or dtype == data.dtype)
        ):
            if copy:
                return data.copy()
            else:
                return data._view()

        # - Cases checked above all return/raise before reaching here - #

        tdarr = TimedeltaArray._from_sequence_not_strict(
            data, freq=freq, unit=unit, dtype=dtype, copy=copy
        )
        refs = None
        if not copy and isinstance(data, (ABCSeries, Index)):
            refs = data._references

        return cls._simple_new(tdarr, name=name, refs=refs)

    # -------------------------------------------------------------------

    def _is_comparable_dtype(self, dtype: DtypeObj) -> bool:
        """
        Can we compare values of the given dtype to our own?
        """
        return lib.is_np_dtype(dtype, "m")  # aka self._data._is_recognized_dtype

    # -------------------------------------------------------------------
    # Indexing Methods

    def get_loc(self, key):
        """
        Get integer location for requested label

        Returns
        -------
        loc : int, slice, or ndarray[int]
        """
        self._check_indexing_error(key)

        try:
            key = self._data._validate_scalar(key, unbox=False)
        except TypeError as err:
            raise KeyError(key) from err

        return Index.get_loc(self, key)

    def _parse_with_reso(self, label: str):
        # the "with_reso" is a no-op for TimedeltaIndex
        parsed = Timedelta(label)
        return parsed, None

    def _parsed_string_to_bounds(self, reso, parsed: Timedelta):
        # reso is unused, included to match signature of DTI/PI
        lbound = parsed.round(parsed.resolution_string)
        rbound = lbound + to_offset(parsed.resolution_string) - Timedelta(1, "ns")
        return lbound, rbound

    # -------------------------------------------------------------------

    @property
    def inferred_type(self) -> str:
        return "timedelta64"


def timedelta_range(
    start=None,
    end=None,
    periods: int | None = None,
    freq=None,
    name=None,
    closed=None,
    *,
    unit: str | None = None,
) -> TimedeltaIndex:
    """
    Return a fixed frequency TimedeltaIndex with day as the default.

    Parameters
    ----------
    start : str or timedelta-like, default None
        Left bound for generating timedeltas.
    end : str or timedelta-like, default None
        Right bound for generating timedeltas.
    periods : int, default None
        Number of periods to generate.
    freq : str, Timedelta, datetime.timedelta, or DateOffset, default 'D'
        Frequency strings can have multiples, e.g. '5h'.
    name : str, default None
        Name of the resulting TimedeltaIndex.
    closed : str, default None
        Make the interval closed with respect to the given frequency to
        the 'left', 'right', or both sides (None).
    unit : str, default None
        Specify the desired resolution of the result.

        .. versionadded:: 2.0.0

    Returns
    -------
    TimedeltaIndex

    Notes
    -----
    Of the four parameters ``start``, ``end``, ``periods``, and ``freq``,
    exactly three must be specified. If ``freq`` is omitted, the resulting
    ``TimedeltaIndex`` will have ``periods`` linearly spaced elements between
    ``start`` and ``end`` (closed on both sides).

    To learn more about the frequency strings, please see `this link
    <https://pandas.pydata.org/pandas-docs/stable/user_guide/timeseries.html#offset-aliases>`__.

    Examples
    --------
    >>> pd.timedelta_range(start='1 day', periods=4)
    TimedeltaIndex(['1 days', '2 days', '3 days', '4 days'],
                   dtype='timedelta64[ns]', freq='D')

    The ``closed`` parameter specifies which endpoint is included.  The default
    behavior is to include both endpoints.

    >>> pd.timedelta_range(start='1 day', periods=4, closed='right')
    TimedeltaIndex(['2 days', '3 days', '4 days'],
                   dtype='timedelta64[ns]', freq='D')

    The ``freq`` parameter specifies the frequency of the TimedeltaIndex.
    Only fixed frequencies can be passed, non-fixed frequencies such as
    'M' (month end) will raise.

    >>> pd.timedelta_range(start='1 day', end='2 days', freq='6h')
    TimedeltaIndex(['1 days 00:00:00', '1 days 06:00:00', '1 days 12:00:00',
                    '1 days 18:00:00', '2 days 00:00:00'],
                   dtype='timedelta64[ns]', freq='6h')

    Specify ``start``, ``end``, and ``periods``; the frequency is generated
    automatically (linearly spaced).

    >>> pd.timedelta_range(start='1 day', end='5 days', periods=4)
    TimedeltaIndex(['1 days 00:00:00', '2 days 08:00:00', '3 days 16:00:00',
                    '5 days 00:00:00'],
                   dtype='timedelta64[ns]', freq=None)

    **Specify a unit**

    >>> pd.timedelta_range("1 Day", periods=3, freq="100000D", unit="s")
    TimedeltaIndex(['1 days', '100001 days', '200001 days'],
                   dtype='timedelta64[s]', freq='100000D')
    """
    if freq is None and com.any_none(periods, start, end):
        freq = "D"

    freq = to_offset(freq)
    tdarr = TimedeltaArray._generate_range(
        start, end, periods, freq, closed=closed, unit=unit
    )
    return TimedeltaIndex._simple_new(tdarr, name=name)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\agca\extensions.py ===
"""Finite extensions of ring domains."""

from sympy.polys.domains.domain import Domain
from sympy.polys.domains.domainelement import DomainElement
from sympy.polys.polyerrors import (CoercionFailed, NotInvertible,
        GeneratorsError)
from sympy.polys.polytools import Poly
from sympy.printing.defaults import DefaultPrinting


class ExtensionElement(DomainElement, DefaultPrinting):
    """
    Element of a finite extension.

    A class of univariate polynomials modulo the ``modulus``
    of the extension ``ext``. It is represented by the
    unique polynomial ``rep`` of lowest degree. Both
    ``rep`` and the representation ``mod`` of ``modulus``
    are of class DMP.

    """
    __slots__ = ('rep', 'ext')

    def __init__(self, rep, ext):
        self.rep = rep
        self.ext = ext

    def parent(f):
        return f.ext

    def as_expr(f):
        return f.ext.to_sympy(f)

    def __bool__(f):
        return bool(f.rep)

    def __pos__(f):
        return f

    def __neg__(f):
        return ExtElem(-f.rep, f.ext)

    def _get_rep(f, g):
        if isinstance(g, ExtElem):
            if g.ext == f.ext:
                return g.rep
            else:
                return None
        else:
            try:
                g = f.ext.convert(g)
                return g.rep
            except CoercionFailed:
                return None

    def __add__(f, g):
        rep = f._get_rep(g)
        if rep is not None:
            return ExtElem(f.rep + rep, f.ext)
        else:
            return NotImplemented

    __radd__ = __add__

    def __sub__(f, g):
        rep = f._get_rep(g)
        if rep is not None:
            return ExtElem(f.rep - rep, f.ext)
        else:
            return NotImplemented

    def __rsub__(f, g):
        rep = f._get_rep(g)
        if rep is not None:
            return ExtElem(rep - f.rep, f.ext)
        else:
            return NotImplemented

    def __mul__(f, g):
        rep = f._get_rep(g)
        if rep is not None:
            return ExtElem((f.rep * rep) % f.ext.mod, f.ext)
        else:
            return NotImplemented

    __rmul__ = __mul__

    def _divcheck(f):
        """Raise if division is not implemented for this divisor"""
        if not f:
            raise NotInvertible('Zero divisor')
        elif f.ext.is_Field:
            return True
        elif f.rep.is_ground and f.ext.domain.is_unit(f.rep.LC()):
            return True
        else:
            # Some cases like (2*x + 2)/2 over ZZ will fail here. It is
            # unclear how to implement division in general if the ground
            # domain is not a field so for now it was decided to restrict the
            # implementation to division by invertible constants.
            msg = (f"Can not invert {f} in {f.ext}. "
                    "Only division by invertible constants is implemented.")
            raise NotImplementedError(msg)

    def inverse(f):
        """Multiplicative inverse.

        Raises
        ======

        NotInvertible
            If the element is a zero divisor.

        """
        f._divcheck()

        if f.ext.is_Field:
            invrep = f.rep.invert(f.ext.mod)
        else:
            R = f.ext.ring
            invrep = R.exquo(R.one, f.rep)

        return ExtElem(invrep, f.ext)

    def __truediv__(f, g):
        rep = f._get_rep(g)
        if rep is None:
            return NotImplemented
        g = ExtElem(rep, f.ext)

        try:
            ginv = g.inverse()
        except NotInvertible:
            raise ZeroDivisionError(f"{f} / {g}")

        return f * ginv

    __floordiv__ = __truediv__

    def __rtruediv__(f, g):
        try:
            g = f.ext.convert(g)
        except CoercionFailed:
            return NotImplemented
        return g / f

    __rfloordiv__ = __rtruediv__

    def __mod__(f, g):
        rep = f._get_rep(g)
        if rep is None:
            return NotImplemented
        g = ExtElem(rep, f.ext)

        try:
            g._divcheck()
        except NotInvertible:
            raise ZeroDivisionError(f"{f} % {g}")

        # Division where defined is always exact so there is no remainder
        return f.ext.zero

    def __rmod__(f, g):
        try:
            g = f.ext.convert(g)
        except CoercionFailed:
            return NotImplemented
        return g % f

    def __pow__(f, n):
        if not isinstance(n, int):
            raise TypeError("exponent of type 'int' expected")
        if n < 0:
            try:
                f, n = f.inverse(), -n
            except NotImplementedError:
                raise ValueError("negative powers are not defined")

        b = f.rep
        m = f.ext.mod
        r = f.ext.one.rep
        while n > 0:
            if n % 2:
                r = (r*b) % m
            b = (b*b) % m
            n //= 2

        return ExtElem(r, f.ext)

    def __eq__(f, g):
        if isinstance(g, ExtElem):
            return f.rep == g.rep and f.ext == g.ext
        else:
            return NotImplemented

    def __ne__(f, g):
        return not f == g

    def __hash__(f):
        return hash((f.rep, f.ext))

    def __str__(f):
        from sympy.printing.str import sstr
        return sstr(f.as_expr())

    __repr__ = __str__

    @property
    def is_ground(f):
        return f.rep.is_ground

    def to_ground(f):
        [c] = f.rep.to_list()
        return c

ExtElem = ExtensionElement


class MonogenicFiniteExtension(Domain):
    r"""
    Finite extension generated by an integral element.

    The generator is defined by a monic univariate
    polynomial derived from the argument ``mod``.

    A shorter alias is ``FiniteExtension``.

    Examples
    ========

    Quadratic integer ring $\mathbb{Z}[\sqrt2]$:

    >>> from sympy import Symbol, Poly
    >>> from sympy.polys.agca.extensions import FiniteExtension
    >>> x = Symbol('x')
    >>> R = FiniteExtension(Poly(x**2 - 2)); R
    ZZ[x]/(x**2 - 2)
    >>> R.rank
    2
    >>> R(1 + x)*(3 - 2*x)
    x - 1

    Finite field $GF(5^3)$ defined by the primitive
    polynomial $x^3 + x^2 + 2$ (over $\mathbb{Z}_5$).

    >>> F = FiniteExtension(Poly(x**3 + x**2 + 2, modulus=5)); F
    GF(5)[x]/(x**3 + x**2 + 2)
    >>> F.basis
    (1, x, x**2)
    >>> F(x + 3)/(x**2 + 2)
    -2*x**2 + x + 2

    Function field of an elliptic curve:

    >>> t = Symbol('t')
    >>> FiniteExtension(Poly(t**2 - x**3 - x + 1, t, field=True))
    ZZ(x)[t]/(t**2 - x**3 - x + 1)

    """
    is_FiniteExtension = True

    dtype = ExtensionElement

    def __init__(self, mod):
        if not (isinstance(mod, Poly) and mod.is_univariate):
            raise TypeError("modulus must be a univariate Poly")

        # Using auto=True (default) potentially changes the ground domain to a
        # field whereas auto=False raises if division is not exact.  We'll let
        # the caller decide whether or not they want to put the ground domain
        # over a field. In most uses mod is already monic.
        mod = mod.monic(auto=False)

        self.rank = mod.degree()
        self.modulus = mod
        self.mod = mod.rep  # DMP representation

        self.domain = dom = mod.domain
        self.ring = dom.old_poly_ring(*mod.gens)

        self.zero = self.convert(self.ring.zero)
        self.one = self.convert(self.ring.one)

        gen = self.ring.gens[0]
        self.symbol = self.ring.symbols[0]
        self.generator = self.convert(gen)
        self.basis = tuple(self.convert(gen**i) for i in range(self.rank))

        # XXX: It might be necessary to check mod.is_irreducible here
        self.is_Field = self.domain.is_Field

    def new(self, arg):
        rep = self.ring.convert(arg)
        return ExtElem(rep % self.mod, self)

    def __eq__(self, other):
        if not isinstance(other, FiniteExtension):
            return False
        return self.modulus == other.modulus

    def __hash__(self):
        return hash((self.__class__.__name__, self.modulus))

    def __str__(self):
        return "%s/(%s)" % (self.ring, self.modulus.as_expr())

    __repr__ = __str__

    @property
    def has_CharacteristicZero(self):
        return self.domain.has_CharacteristicZero

    def characteristic(self):
        return self.domain.characteristic()

    def convert(self, f, base=None):
        rep = self.ring.convert(f, base)
        return ExtElem(rep % self.mod, self)

    def convert_from(self, f, base):
        rep = self.ring.convert(f, base)
        return ExtElem(rep % self.mod, self)

    def to_sympy(self, f):
        return self.ring.to_sympy(f.rep)

    def from_sympy(self, f):
        return self.convert(f)

    def set_domain(self, K):
        mod = self.modulus.set_domain(K)
        return self.__class__(mod)

    def drop(self, *symbols):
        if self.symbol in symbols:
            raise GeneratorsError('Can not drop generator from FiniteExtension')
        K = self.domain.drop(*symbols)
        return self.set_domain(K)

    def quo(self, f, g):
        return self.exquo(f, g)

    def exquo(self, f, g):
        rep = self.ring.exquo(f.rep, g.rep)
        return ExtElem(rep % self.mod, self)

    def is_negative(self, a):
        return False

    def is_unit(self, a):
        if self.is_Field:
            return bool(a)
        elif a.is_ground:
            return self.domain.is_unit(a.to_ground())

FiniteExtension = MonogenicFiniteExtension

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\urllib3\http2\connection.py ===
from __future__ import annotations

import logging
import re
import threading
import types
import typing

import h2.config  # type: ignore[import-untyped]
import h2.connection  # type: ignore[import-untyped]
import h2.events  # type: ignore[import-untyped]

from .._base_connection import _TYPE_BODY
from .._collections import HTTPHeaderDict
from ..connection import HTTPSConnection, _get_default_user_agent
from ..exceptions import ConnectionError
from ..response import BaseHTTPResponse

orig_HTTPSConnection = HTTPSConnection

T = typing.TypeVar("T")

log = logging.getLogger(__name__)

RE_IS_LEGAL_HEADER_NAME = re.compile(rb"^[!#$%&'*+\-.^_`|~0-9a-z]+$")
RE_IS_ILLEGAL_HEADER_VALUE = re.compile(rb"[\0\x00\x0a\x0d\r\n]|^[ \r\n\t]|[ \r\n\t]$")


def _is_legal_header_name(name: bytes) -> bool:
    """
    "An implementation that validates fields according to the definitions in Sections
    5.1 and 5.5 of [HTTP] only needs an additional check that field names do not
    include uppercase characters." (https://httpwg.org/specs/rfc9113.html#n-field-validity)

    `http.client._is_legal_header_name` does not validate the field name according to the
    HTTP 1.1 spec, so we do that here, in addition to checking for uppercase characters.

    This does not allow for the `:` character in the header name, so should not
    be used to validate pseudo-headers.
    """
    return bool(RE_IS_LEGAL_HEADER_NAME.match(name))


def _is_illegal_header_value(value: bytes) -> bool:
    """
    "A field value MUST NOT contain the zero value (ASCII NUL, 0x00), line feed
    (ASCII LF, 0x0a), or carriage return (ASCII CR, 0x0d) at any position. A field
    value MUST NOT start or end with an ASCII whitespace character (ASCII SP or HTAB,
    0x20 or 0x09)." (https://httpwg.org/specs/rfc9113.html#n-field-validity)
    """
    return bool(RE_IS_ILLEGAL_HEADER_VALUE.search(value))


class _LockedObject(typing.Generic[T]):
    """
    A wrapper class that hides a specific object behind a lock.
    The goal here is to provide a simple way to protect access to an object
    that cannot safely be simultaneously accessed from multiple threads. The
    intended use of this class is simple: take hold of it with a context
    manager, which returns the protected object.
    """

    __slots__ = (
        "lock",
        "_obj",
    )

    def __init__(self, obj: T):
        self.lock = threading.RLock()
        self._obj = obj

    def __enter__(self) -> T:
        self.lock.acquire()
        return self._obj

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: types.TracebackType | None,
    ) -> None:
        self.lock.release()


class HTTP2Connection(HTTPSConnection):
    def __init__(
        self, host: str, port: int | None = None, **kwargs: typing.Any
    ) -> None:
        self._h2_conn = self._new_h2_conn()
        self._h2_stream: int | None = None
        self._headers: list[tuple[bytes, bytes]] = []

        if "proxy" in kwargs or "proxy_config" in kwargs:  # Defensive:
            raise NotImplementedError("Proxies aren't supported with HTTP/2")

        super().__init__(host, port, **kwargs)

        if self._tunnel_host is not None:
            raise NotImplementedError("Tunneling isn't supported with HTTP/2")

    def _new_h2_conn(self) -> _LockedObject[h2.connection.H2Connection]:
        config = h2.config.H2Configuration(client_side=True)
        return _LockedObject(h2.connection.H2Connection(config=config))

    def connect(self) -> None:
        super().connect()
        with self._h2_conn as conn:
            conn.initiate_connection()
            if data_to_send := conn.data_to_send():
                self.sock.sendall(data_to_send)

    def putrequest(  # type: ignore[override]
        self,
        method: str,
        url: str,
        **kwargs: typing.Any,
    ) -> None:
        """putrequest
        This deviates from the HTTPConnection method signature since we never need to override
        sending accept-encoding headers or the host header.
        """
        if "skip_host" in kwargs:
            raise NotImplementedError("`skip_host` isn't supported")
        if "skip_accept_encoding" in kwargs:
            raise NotImplementedError("`skip_accept_encoding` isn't supported")

        self._request_url = url or "/"
        self._validate_path(url)  # type: ignore[attr-defined]

        if ":" in self.host:
            authority = f"[{self.host}]:{self.port or 443}"
        else:
            authority = f"{self.host}:{self.port or 443}"

        self._headers.append((b":scheme", b"https"))
        self._headers.append((b":method", method.encode()))
        self._headers.append((b":authority", authority.encode()))
        self._headers.append((b":path", url.encode()))

        with self._h2_conn as conn:
            self._h2_stream = conn.get_next_available_stream_id()

    def putheader(self, header: str | bytes, *values: str | bytes) -> None:  # type: ignore[override]
        # TODO SKIPPABLE_HEADERS from urllib3 are ignored.
        header = header.encode() if isinstance(header, str) else header
        header = header.lower()  # A lot of upstream code uses capitalized headers.
        if not _is_legal_header_name(header):
            raise ValueError(f"Illegal header name {str(header)}")

        for value in values:
            value = value.encode() if isinstance(value, str) else value
            if _is_illegal_header_value(value):
                raise ValueError(f"Illegal header value {str(value)}")
            self._headers.append((header, value))

    def endheaders(self, message_body: typing.Any = None) -> None:  # type: ignore[override]
        if self._h2_stream is None:
            raise ConnectionError("Must call `putrequest` first.")

        with self._h2_conn as conn:
            conn.send_headers(
                stream_id=self._h2_stream,
                headers=self._headers,
                end_stream=(message_body is None),
            )
            if data_to_send := conn.data_to_send():
                self.sock.sendall(data_to_send)
        self._headers = []  # Reset headers for the next request.

    def send(self, data: typing.Any) -> None:
        """Send data to the server.
        `data` can be: `str`, `bytes`, an iterable, or file-like objects
        that support a .read() method.
        """
        if self._h2_stream is None:
            raise ConnectionError("Must call `putrequest` first.")

        with self._h2_conn as conn:
            if data_to_send := conn.data_to_send():
                self.sock.sendall(data_to_send)

            if hasattr(data, "read"):  # file-like objects
                while True:
                    chunk = data.read(self.blocksize)
                    if not chunk:
                        break
                    if isinstance(chunk, str):
                        chunk = chunk.encode()  # pragma: no cover
                    conn.send_data(self._h2_stream, chunk, end_stream=False)
                    if data_to_send := conn.data_to_send():
                        self.sock.sendall(data_to_send)
                conn.end_stream(self._h2_stream)
                return

            if isinstance(data, str):  # str -> bytes
                data = data.encode()

            try:
                if isinstance(data, bytes):
                    conn.send_data(self._h2_stream, data, end_stream=True)
                    if data_to_send := conn.data_to_send():
                        self.sock.sendall(data_to_send)
                else:
                    for chunk in data:
                        conn.send_data(self._h2_stream, chunk, end_stream=False)
                        if data_to_send := conn.data_to_send():
                            self.sock.sendall(data_to_send)
                    conn.end_stream(self._h2_stream)
            except TypeError:
                raise TypeError(
                    "`data` should be str, bytes, iterable, or file. got %r"
                    % type(data)
                )

    def set_tunnel(
        self,
        host: str,
        port: int | None = None,
        headers: typing.Mapping[str, str] | None = None,
        scheme: str = "http",
    ) -> None:
        raise NotImplementedError(
            "HTTP/2 does not support setting up a tunnel through a proxy"
        )

    def getresponse(  # type: ignore[override]
        self,
    ) -> HTTP2Response:
        status = None
        data = bytearray()
        with self._h2_conn as conn:
            end_stream = False
            while not end_stream:
                # TODO: Arbitrary read value.
                if received_data := self.sock.recv(65535):
                    events = conn.receive_data(received_data)
                    for event in events:
                        if isinstance(event, h2.events.ResponseReceived):
                            headers = HTTPHeaderDict()
                            for header, value in event.headers:
                                if header == b":status":
                                    status = int(value.decode())
                                else:
                                    headers.add(
                                        header.decode("ascii"), value.decode("ascii")
                                    )

                        elif isinstance(event, h2.events.DataReceived):
                            data += event.data
                            conn.acknowledge_received_data(
                                event.flow_controlled_length, event.stream_id
                            )

                        elif isinstance(event, h2.events.StreamEnded):
                            end_stream = True

                if data_to_send := conn.data_to_send():
                    self.sock.sendall(data_to_send)

        assert status is not None
        return HTTP2Response(
            status=status,
            headers=headers,
            request_url=self._request_url,
            data=bytes(data),
        )

    def request(  # type: ignore[override]
        self,
        method: str,
        url: str,
        body: _TYPE_BODY | None = None,
        headers: typing.Mapping[str, str] | None = None,
        *,
        preload_content: bool = True,
        decode_content: bool = True,
        enforce_content_length: bool = True,
        **kwargs: typing.Any,
    ) -> None:
        """Send an HTTP/2 request"""
        if "chunked" in kwargs:
            # TODO this is often present from upstream.
            # raise NotImplementedError("`chunked` isn't supported with HTTP/2")
            pass

        if self.sock is not None:
            self.sock.settimeout(self.timeout)

        self.putrequest(method, url)

        headers = headers or {}
        for k, v in headers.items():
            if k.lower() == "transfer-encoding" and v == "chunked":
                continue
            else:
                self.putheader(k, v)

        if b"user-agent" not in dict(self._headers):
            self.putheader(b"user-agent", _get_default_user_agent())

        if body:
            self.endheaders(message_body=body)
            self.send(body)
        else:
            self.endheaders()

    def close(self) -> None:
        with self._h2_conn as conn:
            try:
                conn.close_connection()
                if data := conn.data_to_send():
                    self.sock.sendall(data)
            except Exception:
                pass

        # Reset all our HTTP/2 connection state.
        self._h2_conn = self._new_h2_conn()
        self._h2_stream = None
        self._headers = []

        super().close()


class HTTP2Response(BaseHTTPResponse):
    # TODO: This is a woefully incomplete response object, but works for non-streaming.
    def __init__(
        self,
        status: int,
        headers: HTTPHeaderDict,
        request_url: str,
        data: bytes,
        decode_content: bool = False,  # TODO: support decoding
    ) -> None:
        super().__init__(
            status=status,
            headers=headers,
            # Following CPython, we map HTTP versions to major * 10 + minor integers
            version=20,
            version_string="HTTP/2",
            # No reason phrase in HTTP/2
            reason=None,
            decode_content=decode_content,
            request_url=request_url,
        )
        self._data = data
        self.length_remaining = 0

    @property
    def data(self) -> bytes:
        return self._data

    def get_redirect_location(self) -> None:
        return None

    def close(self) -> None:
        pass

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\_core\_machar.py ===
"""
Machine arithmetic - determine the parameters of the
floating-point arithmetic system

Author: Pearu Peterson, September 2003

"""
__all__ = ['MachAr']

from ._ufunc_config import errstate
from .fromnumeric import any

# Need to speed this up...especially for longdouble

# Deprecated 2021-10-20, NumPy 1.22
class MachAr:
    """
    Diagnosing machine parameters.

    Attributes
    ----------
    ibeta : int
        Radix in which numbers are represented.
    it : int
        Number of base-`ibeta` digits in the floating point mantissa M.
    machep : int
        Exponent of the smallest (most negative) power of `ibeta` that,
        added to 1.0, gives something different from 1.0
    eps : float
        Floating-point number ``beta**machep`` (floating point precision)
    negep : int
        Exponent of the smallest power of `ibeta` that, subtracted
        from 1.0, gives something different from 1.0.
    epsneg : float
        Floating-point number ``beta**negep``.
    iexp : int
        Number of bits in the exponent (including its sign and bias).
    minexp : int
        Smallest (most negative) power of `ibeta` consistent with there
        being no leading zeros in the mantissa.
    xmin : float
        Floating-point number ``beta**minexp`` (the smallest [in
        magnitude] positive floating point number with full precision).
    maxexp : int
        Smallest (positive) power of `ibeta` that causes overflow.
    xmax : float
        ``(1-epsneg) * beta**maxexp`` (the largest [in magnitude]
        usable floating value).
    irnd : int
        In ``range(6)``, information on what kind of rounding is done
        in addition, and on how underflow is handled.
    ngrd : int
        Number of 'guard digits' used when truncating the product
        of two mantissas to fit the representation.
    epsilon : float
        Same as `eps`.
    tiny : float
        An alias for `smallest_normal`, kept for backwards compatibility.
    huge : float
        Same as `xmax`.
    precision : float
        ``- int(-log10(eps))``
    resolution : float
        ``- 10**(-precision)``
    smallest_normal : float
        The smallest positive floating point number with 1 as leading bit in
        the mantissa following IEEE-754. Same as `xmin`.
    smallest_subnormal : float
        The smallest positive floating point number with 0 as leading bit in
        the mantissa following IEEE-754.

    Parameters
    ----------
    float_conv : function, optional
        Function that converts an integer or integer array to a float
        or float array. Default is `float`.
    int_conv : function, optional
        Function that converts a float or float array to an integer or
        integer array. Default is `int`.
    float_to_float : function, optional
        Function that converts a float array to float. Default is `float`.
        Note that this does not seem to do anything useful in the current
        implementation.
    float_to_str : function, optional
        Function that converts a single float to a string. Default is
        ``lambda v:'%24.16e' %v``.
    title : str, optional
        Title that is printed in the string representation of `MachAr`.

    See Also
    --------
    finfo : Machine limits for floating point types.
    iinfo : Machine limits for integer types.

    References
    ----------
    .. [1] Press, Teukolsky, Vetterling and Flannery,
           "Numerical Recipes in C++," 2nd ed,
           Cambridge University Press, 2002, p. 31.

    """

    def __init__(self, float_conv=float, int_conv=int,
                 float_to_float=float,
                 float_to_str=lambda v: f'{v:24.16e}',
                 title='Python floating point number'):
        """

        float_conv - convert integer to float (array)
        int_conv   - convert float (array) to integer
        float_to_float - convert float array to float
        float_to_str - convert array float to str
        title        - description of used floating point numbers

        """
        # We ignore all errors here because we are purposely triggering
        # underflow to detect the properties of the running arch.
        with errstate(under='ignore'):
            self._do_init(float_conv, int_conv, float_to_float, float_to_str, title)

    def _do_init(self, float_conv, int_conv, float_to_float, float_to_str, title):
        max_iterN = 10000
        msg = "Did not converge after %d tries with %s"
        one = float_conv(1)
        two = one + one
        zero = one - one

        # Do we really need to do this?  Aren't they 2 and 2.0?
        # Determine ibeta and beta
        a = one
        for _ in range(max_iterN):
            a = a + a
            temp = a + one
            temp1 = temp - a
            if any(temp1 - one != zero):
                break
        else:
            raise RuntimeError(msg % (_, one.dtype))
        b = one
        for _ in range(max_iterN):
            b = b + b
            temp = a + b
            itemp = int_conv(temp - a)
            if any(itemp != 0):
                break
        else:
            raise RuntimeError(msg % (_, one.dtype))
        ibeta = itemp
        beta = float_conv(ibeta)

        # Determine it and irnd
        it = -1
        b = one
        for _ in range(max_iterN):
            it = it + 1
            b = b * beta
            temp = b + one
            temp1 = temp - b
            if any(temp1 - one != zero):
                break
        else:
            raise RuntimeError(msg % (_, one.dtype))

        betah = beta / two
        a = one
        for _ in range(max_iterN):
            a = a + a
            temp = a + one
            temp1 = temp - a
            if any(temp1 - one != zero):
                break
        else:
            raise RuntimeError(msg % (_, one.dtype))
        temp = a + betah
        irnd = 0
        if any(temp - a != zero):
            irnd = 1
        tempa = a + beta
        temp = tempa + betah
        if irnd == 0 and any(temp - tempa != zero):
            irnd = 2

        # Determine negep and epsneg
        negep = it + 3
        betain = one / beta
        a = one
        for i in range(negep):
            a = a * betain
        b = a
        for _ in range(max_iterN):
            temp = one - a
            if any(temp - one != zero):
                break
            a = a * beta
            negep = negep - 1
            # Prevent infinite loop on PPC with gcc 4.0:
            if negep < 0:
                raise RuntimeError("could not determine machine tolerance "
                                   "for 'negep', locals() -> %s" % (locals()))
        else:
            raise RuntimeError(msg % (_, one.dtype))
        negep = -negep
        epsneg = a

        # Determine machep and eps
        machep = - it - 3
        a = b

        for _ in range(max_iterN):
            temp = one + a
            if any(temp - one != zero):
                break
            a = a * beta
            machep = machep + 1
        else:
            raise RuntimeError(msg % (_, one.dtype))
        eps = a

        # Determine ngrd
        ngrd = 0
        temp = one + eps
        if irnd == 0 and any(temp * one - one != zero):
            ngrd = 1

        # Determine iexp
        i = 0
        k = 1
        z = betain
        t = one + eps
        nxres = 0
        for _ in range(max_iterN):
            y = z
            z = y * y
            a = z * one  # Check here for underflow
            temp = z * t
            if any(a + a == zero) or any(abs(z) >= y):
                break
            temp1 = temp * betain
            if any(temp1 * beta == z):
                break
            i = i + 1
            k = k + k
        else:
            raise RuntimeError(msg % (_, one.dtype))
        if ibeta != 10:
            iexp = i + 1
            mx = k + k
        else:
            iexp = 2
            iz = ibeta
            while k >= iz:
                iz = iz * ibeta
                iexp = iexp + 1
            mx = iz + iz - 1

        # Determine minexp and xmin
        for _ in range(max_iterN):
            xmin = y
            y = y * betain
            a = y * one
            temp = y * t
            if any((a + a) != zero) and any(abs(y) < xmin):
                k = k + 1
                temp1 = temp * betain
                if any(temp1 * beta == y) and any(temp != y):
                    nxres = 3
                    xmin = y
                    break
            else:
                break
        else:
            raise RuntimeError(msg % (_, one.dtype))
        minexp = -k

        # Determine maxexp, xmax
        if mx <= k + k - 3 and ibeta != 10:
            mx = mx + mx
            iexp = iexp + 1
        maxexp = mx + minexp
        irnd = irnd + nxres
        if irnd >= 2:
            maxexp = maxexp - 2
        i = maxexp + minexp
        if ibeta == 2 and not i:
            maxexp = maxexp - 1
        if i > 20:
            maxexp = maxexp - 1
        if any(a != y):
            maxexp = maxexp - 2
        xmax = one - epsneg
        if any(xmax * one != xmax):
            xmax = one - beta * epsneg
        xmax = xmax / (xmin * beta * beta * beta)
        i = maxexp + minexp + 3
        for j in range(i):
            if ibeta == 2:
                xmax = xmax + xmax
            else:
                xmax = xmax * beta

        smallest_subnormal = abs(xmin / beta ** (it))

        self.ibeta = ibeta
        self.it = it
        self.negep = negep
        self.epsneg = float_to_float(epsneg)
        self._str_epsneg = float_to_str(epsneg)
        self.machep = machep
        self.eps = float_to_float(eps)
        self._str_eps = float_to_str(eps)
        self.ngrd = ngrd
        self.iexp = iexp
        self.minexp = minexp
        self.xmin = float_to_float(xmin)
        self._str_xmin = float_to_str(xmin)
        self.maxexp = maxexp
        self.xmax = float_to_float(xmax)
        self._str_xmax = float_to_str(xmax)
        self.irnd = irnd

        self.title = title
        # Commonly used parameters
        self.epsilon = self.eps
        self.tiny = self.xmin
        self.huge = self.xmax
        self.smallest_normal = self.xmin
        self._str_smallest_normal = float_to_str(self.xmin)
        self.smallest_subnormal = float_to_float(smallest_subnormal)
        self._str_smallest_subnormal = float_to_str(smallest_subnormal)

        import math
        self.precision = int(-math.log10(float_to_float(self.eps)))
        ten = two + two + two + two + two
        resolution = ten ** (-self.precision)
        self.resolution = float_to_float(resolution)
        self._str_resolution = float_to_str(resolution)

    def __str__(self):
        fmt = (
           'Machine parameters for %(title)s\n'
           '---------------------------------------------------------------------\n'
           'ibeta=%(ibeta)s it=%(it)s iexp=%(iexp)s ngrd=%(ngrd)s irnd=%(irnd)s\n'
           'machep=%(machep)s     eps=%(_str_eps)s (beta**machep == epsilon)\n'
           'negep =%(negep)s  epsneg=%(_str_epsneg)s (beta**epsneg)\n'
           'minexp=%(minexp)s   xmin=%(_str_xmin)s (beta**minexp == tiny)\n'
           'maxexp=%(maxexp)s    xmax=%(_str_xmax)s ((1-epsneg)*beta**maxexp == huge)\n'
           'smallest_normal=%(smallest_normal)s    '
           'smallest_subnormal=%(smallest_subnormal)s\n'
           '---------------------------------------------------------------------\n'
           )
        return fmt % self.__dict__


if __name__ == '__main__':
    print(MachAr())

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_gpt_attention_megatron.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
from logging import getLogger

import numpy as np
from fusion_gpt_attention import FusionGptAttentionPastBase
from onnx import helper
from onnx_model import OnnxModel

logger = getLogger(__name__)


def is_close(value, expected_value):
    return abs(value - expected_value) <= 1e-6


class FusionGptAttentionMegatron(FusionGptAttentionPastBase):
    """
    Fuse GPT-2 Attention with past state subgraph from Megatron into one Attention node.
    """

    def __init__(self, model: OnnxModel, num_heads: int):
        super().__init__(model, num_heads)

    def fuse_attention_node(
        self,
        matmul_before_split,
        add_before_split,
        past,
        present,
        input,
        reshape_qkv,
        mask,
    ):
        attention_node_name = self.model.create_node_name("GptAttention")
        int32_mask = self.cast_attention_mask(mask)
        output = reshape_qkv.output[0]
        i = 1 if (add_before_split.input[0] == matmul_before_split.output[0]) else 0
        attention_node = helper.make_node(
            "Attention",
            inputs=[
                input,
                matmul_before_split.input[1],
                add_before_split.input[i],
                int32_mask,
                past,
            ],
            outputs=[output, present],
            name=attention_node_name,
        )
        attention_node.domain = "com.microsoft"
        attention_node.attribute.extend(
            [
                helper.make_attribute("num_heads", self.num_heads),
                helper.make_attribute("unidirectional", 0),  # unidirectional shall not be ON for 4D attention mask
            ]
        )
        if self.mask_filter_value is not None:
            attention_node.attribute.extend([helper.make_attribute("mask_filter_value", float(self.mask_filter_value))])

        nodes_to_add = [attention_node]
        self.nodes_to_add.extend(nodes_to_add)

        for node in nodes_to_add:
            self.node_name_to_graph_name[node.name] = self.this_graph_name

        self.nodes_to_remove.append(reshape_qkv)

        # we rely on prune_graph() to clean old subgraph nodes
        self.prune_graph = True

    def match_mask(self, sub_qk, mul_qk, matmul_qk, layernorm_before_attention):
        mask_nodes = self.model.match_parent_path(sub_qk, ["Mul", "Sub", "Slice", "Slice"], [1, 0, 1, 0])
        if mask_nodes is None:
            logger.debug("fuse_attention: failed to match unidirectional mask path")
            return None
        (mul_mask, sub_mask, last_slice_mask, slice_mask) = mask_nodes

        if len(mask_nodes) > 1 and mask_nodes[0].op_type == "Mul":
            _, mul_val = self.model.get_constant_input(mask_nodes[0])
            if mul_val != 10000:
                self.mask_filter_value = -mul_val

        if mul_qk.input[1] != last_slice_mask.output[0]:
            logger.debug("fuse_attention failed: mul_qk.input[1] != last_slice_mask.output[0]")
            return None

        if not self.utils.check_node_input_value(mul_mask, 1, 10000.0):
            logger.debug("fuse_attention failed: mul_mask input 1 is not constant 10000.0")
            return None

        if not self.utils.check_node_input_value(sub_mask, 0, 1.0):
            logger.debug("fuse_attention failed: sub_mask input 0 is not constant 1.0")
            return None

        if not self.model.find_graph_input(slice_mask.input[0]):
            logger.info("expect slick_mask input 0 to be graph input")
            return None

        if not self.utils.check_node_input_value(last_slice_mask, 1, [0]):
            logger.debug("fuse_attention failed: last_slice_mask input 1 (starts) is not constant [0]")
            return None

        if not self.utils.check_node_input_value(last_slice_mask, 3, [3]):
            logger.debug("fuse_attention failed: last_slice_mask input 3 (axes) is not constant [3]")
            return False

        if not self.utils.check_node_input_value(last_slice_mask, 4, [1]):
            logger.debug("fuse_attention failed: last_slice_mask input 4 (steps) is not constant [1]")
            return False

        if not self.utils.check_node_input_value(slice_mask, 3, [2]):
            logger.debug("fuse_attention failed: slice_mask input 3 (axes) is not constant [2]")
            return None

        if not self.utils.check_node_input_value(slice_mask, 4, [1]):
            logger.debug("fuse_attention failed: slice_mask input 4 (steps) is not constant [1]")
            return None

        last_slice_path = self.model.match_parent_path(
            last_slice_mask, ["Unsqueeze", "Gather", "Shape", "MatMul"], [2, 0, 0, 0]
        )
        if last_slice_path is None or last_slice_path[-1] != matmul_qk:
            logger.debug("fuse_attention: failed to match last slice path")
            return None

        first_slice_path = self.model.match_parent_path(
            slice_mask, ["Unsqueeze", "Gather", "Shape", "MatMul"], [2, 0, 0, 0]
        )
        if first_slice_path is None or first_slice_path[-1] != matmul_qk:
            logger.debug("fuse_attention: failed to match first slice path")
            return None

        first_slice_sub = self.model.match_parent_path(
            slice_mask,
            ["Unsqueeze", "Sub", "Gather", "Shape", "MatMul"],
            [1, 0, 0, 0, 0],
        )
        if first_slice_sub is None or first_slice_sub[-1] != matmul_qk:
            logger.debug("fuse_attention: failed to match last slice sub path")
            return None

        first_slice_sub_1 = self.model.match_parent_path(
            slice_mask,
            ["Unsqueeze", "Sub", "Gather", "Shape", "LayerNormalization"],
            [1, 0, 1, 0, 0],
        )

        if first_slice_sub_1 is None:
            first_slice_sub_1 = self.model.match_parent_path(
                slice_mask,
                ["Unsqueeze", "Sub", "Gather", "Shape", "SkipLayerNormalization"],
                [1, 0, 1, 0, 0],
            )

        if first_slice_sub_1 is None or first_slice_sub_1[-1] != layernorm_before_attention:
            logger.debug("fuse_attention: failed to match last slice sub path 1")
            return None

        return slice_mask.input[0]

    def fuse(self, normalize_node, input_name_to_nodes, output_name_to_node):
        past = None
        present = None

        is_normalize_node_skiplayernorm = normalize_node.op_type == "SkipLayerNormalization"
        qkv_nodes = None

        if not is_normalize_node_skiplayernorm:
            qkv_nodes = self.model.match_parent_path(
                normalize_node,
                ["Add", "Add", "MatMul", "Reshape", "Transpose", "MatMul"],
                [0, 1, None, 0, 0, 0],
                output_name_to_node=output_name_to_node,
            )
        else:
            qkv_nodes = self.model.match_parent_path(
                normalize_node,
                ["Add", "MatMul", "Reshape", "Transpose", "MatMul"],
                [1, None, 0, 0, 0],
                output_name_to_node=output_name_to_node,
            )

        if qkv_nodes is None:
            return

        skip_input = None
        if not is_normalize_node_skiplayernorm:
            (
                add_skip,
                add_after_attention,
                matmul_after_attention,
                reshape_qkv,
                transpose_qkv,
                matmul_qkv,
            ) = qkv_nodes

            skip_input = add_skip.input[0]
        else:
            (
                add_after_attention,
                matmul_after_attention,
                reshape_qkv,
                transpose_qkv,
                matmul_qkv,
            ) = qkv_nodes

            skip_input = normalize_node.input[0]

        v_nodes = self.model.match_parent_path(
            matmul_qkv,
            [
                "Concat",
                "Transpose",
                "Reshape",
                "Split",
                "Add",
                "MatMul",
                "LayerNormalization",
            ],
            [1, 1, 0, 0, 0, None, 0],
        )

        if v_nodes is None:
            v_nodes = self.model.match_parent_path(
                matmul_qkv,
                [
                    "Concat",
                    "Transpose",
                    "Reshape",
                    "Split",
                    "Add",
                    "MatMul",
                    "SkipLayerNormalization",
                ],
                [1, 1, 0, 0, 0, None, 0],
            )

        if v_nodes is None:
            logger.debug("fuse_attention: failed to match v path")
            return
        (
            concat_v,
            transpose_v,
            reshape_v,
            split_v,
            add_before_split,
            matmul_before_split,
            layernorm_before_attention,
        ) = v_nodes

        if (
            layernorm_before_attention.op_type == "LayerNormalization"
            and skip_input != layernorm_before_attention.input[0]
        ):
            logger.debug("fuse_attention: skip_input != layernorm_before_attention.input[0]")
            return

        if (
            layernorm_before_attention.op_type == "SkipLayerNormalization"
            and skip_input != layernorm_before_attention.output[3]
        ):
            logger.debug("fuse_attention: skip_input != layernorm_before_attention.input[0]")
            return

        qk_nodes = self.model.match_parent_path(matmul_qkv, ["Softmax", "Sub", "Mul", "MatMul"], [0, 0, 0, 0])
        if qk_nodes is None:
            logger.debug("fuse_attention: failed to match qk path")
            return None
        (softmax_qk, sub_qk, mul_qk, matmul_qk) = qk_nodes
        if self.model.get_node_attribute(softmax_qk, "axis") != 3:
            logger.debug("fuse_attention failed: softmax_qk axis != 3")
            return None

        attention_mask = self.match_mask(sub_qk, mul_qk, matmul_qk, layernorm_before_attention)

        q_nodes = self.model.match_parent_path(matmul_qk, ["Div", "Transpose", "Reshape", "Split"], [0, 0, 0, 0])
        if q_nodes is None:
            logger.debug("fuse_attention: failed to match q path")
            return
        (div_q, transpose_q, reshape_q, split_q) = q_nodes
        if split_v != split_q:
            logger.debug("fuse_attention: skip since split_v != split_q")
            return

        k_nodes = self.model.match_parent_path(
            matmul_qk,
            ["Div", "Transpose", "Concat", "Transpose", "Reshape", "Split"],
            [1, 0, 0, 1, 0, 0],
        )
        if k_nodes is None:
            logger.debug("fuse_attention: failed to match k path")
            return
        (div_k, _, concat_k, transpose_k, reshape_k, split_k) = k_nodes
        if split_v != split_k:
            logger.debug("fuse_attention: skip since split_v != split_k")
            return

        i, value = self.model.get_constant_input(reshape_k)
        if not (
            isinstance(value, np.ndarray)
            and list(value.shape) == [4]
            and value[0] == 0
            and value[1] == 0
            and value[2] > 0
            and value[3] > 0
        ):
            logger.debug("fuse_attention: reshape constant input is not [0, 0, N, H]")
            return

        num_heads = value[2]
        if num_heads != self.num_heads:
            logger.info(f"Detected num_heads={num_heads}. Ignore user specified value {self.num_heads}")
            self.num_heads = num_heads

        hidden_size_per_head = value[3]
        i, value = self.model.get_constant_input(div_k)
        expected_value = float(np.sqrt(np.sqrt(hidden_size_per_head)))
        if not is_close(value, expected_value):
            logger.debug(f"fuse_attention: div_k value={value} expected={expected_value}")
            return

        i, value = self.model.get_constant_input(div_q)
        if not is_close(value, expected_value):
            logger.debug(f"fuse_attention: div_q value={value} expected={expected_value}")
            return

        # Match past and present paths
        past = self.match_past_pattern_2(concat_k, concat_v, output_name_to_node)
        if past is None:
            logger.debug("fuse_attention: match past failed")
            return
        if not self.model.find_graph_input(past):
            logger.debug("fuse_attention: past is not graph input.")
            # For GPT2LMHeadModel_BeamSearchStep, there is an extra Gather node to select beam index so it is not graph input.

        present = self.match_present(concat_v, input_name_to_nodes)
        if present is None:
            logger.debug("fuse_attention: match present failed")
            return
        if not self.model.find_graph_output(present):
            logger.info("fuse_attention: expect present to be graph output")
            return

        self.fuse_attention_node(
            matmul_before_split,
            add_before_split,
            past,
            present,
            layernorm_before_attention.output[0],
            reshape_qkv,
            attention_mask,
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\computation\scope.py ===
"""
Module for scope operations
"""
from __future__ import annotations

from collections import ChainMap
import datetime
import inspect
from io import StringIO
import itertools
import pprint
import struct
import sys
from typing import TypeVar

import numpy as np

from pandas._libs.tslibs import Timestamp
from pandas.errors import UndefinedVariableError

_KT = TypeVar("_KT")
_VT = TypeVar("_VT")


# https://docs.python.org/3/library/collections.html#chainmap-examples-and-recipes
class DeepChainMap(ChainMap[_KT, _VT]):
    """
    Variant of ChainMap that allows direct updates to inner scopes.

    Only works when all passed mapping are mutable.
    """

    def __setitem__(self, key: _KT, value: _VT) -> None:
        for mapping in self.maps:
            if key in mapping:
                mapping[key] = value
                return
        self.maps[0][key] = value

    def __delitem__(self, key: _KT) -> None:
        """
        Raises
        ------
        KeyError
            If `key` doesn't exist.
        """
        for mapping in self.maps:
            if key in mapping:
                del mapping[key]
                return
        raise KeyError(key)


def ensure_scope(
    level: int, global_dict=None, local_dict=None, resolvers=(), target=None
) -> Scope:
    """Ensure that we are grabbing the correct scope."""
    return Scope(
        level + 1,
        global_dict=global_dict,
        local_dict=local_dict,
        resolvers=resolvers,
        target=target,
    )


def _replacer(x) -> str:
    """
    Replace a number with its hexadecimal representation. Used to tag
    temporary variables with their calling scope's id.
    """
    # get the hex repr of the binary char and remove 0x and pad by pad_size
    # zeros
    try:
        hexin = ord(x)
    except TypeError:
        # bytes literals masquerade as ints when iterating in py3
        hexin = x

    return hex(hexin)


def _raw_hex_id(obj) -> str:
    """Return the padded hexadecimal id of ``obj``."""
    # interpret as a pointer since that's what really what id returns
    packed = struct.pack("@P", id(obj))
    return "".join([_replacer(x) for x in packed])


DEFAULT_GLOBALS = {
    "Timestamp": Timestamp,
    "datetime": datetime.datetime,
    "True": True,
    "False": False,
    "list": list,
    "tuple": tuple,
    "inf": np.inf,
    "Inf": np.inf,
}


def _get_pretty_string(obj) -> str:
    """
    Return a prettier version of obj.

    Parameters
    ----------
    obj : object
        Object to pretty print

    Returns
    -------
    str
        Pretty print object repr
    """
    sio = StringIO()
    pprint.pprint(obj, stream=sio)
    return sio.getvalue()


class Scope:
    """
    Object to hold scope, with a few bells to deal with some custom syntax
    and contexts added by pandas.

    Parameters
    ----------
    level : int
    global_dict : dict or None, optional, default None
    local_dict : dict or Scope or None, optional, default None
    resolvers : list-like or None, optional, default None
    target : object

    Attributes
    ----------
    level : int
    scope : DeepChainMap
    target : object
    temps : dict
    """

    __slots__ = ["level", "scope", "target", "resolvers", "temps"]
    level: int
    scope: DeepChainMap
    resolvers: DeepChainMap
    temps: dict

    def __init__(
        self, level: int, global_dict=None, local_dict=None, resolvers=(), target=None
    ) -> None:
        self.level = level + 1

        # shallow copy because we don't want to keep filling this up with what
        # was there before if there are multiple calls to Scope/_ensure_scope
        self.scope = DeepChainMap(DEFAULT_GLOBALS.copy())
        self.target = target

        if isinstance(local_dict, Scope):
            self.scope.update(local_dict.scope)
            if local_dict.target is not None:
                self.target = local_dict.target
            self._update(local_dict.level)

        frame = sys._getframe(self.level)

        try:
            # shallow copy here because we don't want to replace what's in
            # scope when we align terms (alignment accesses the underlying
            # numpy array of pandas objects)
            scope_global = self.scope.new_child(
                (global_dict if global_dict is not None else frame.f_globals).copy()
            )
            self.scope = DeepChainMap(scope_global)
            if not isinstance(local_dict, Scope):
                scope_local = self.scope.new_child(
                    (local_dict if local_dict is not None else frame.f_locals).copy()
                )
                self.scope = DeepChainMap(scope_local)
        finally:
            del frame

        # assumes that resolvers are going from outermost scope to inner
        if isinstance(local_dict, Scope):
            resolvers += tuple(local_dict.resolvers.maps)
        self.resolvers = DeepChainMap(*resolvers)
        self.temps = {}

    def __repr__(self) -> str:
        scope_keys = _get_pretty_string(list(self.scope.keys()))
        res_keys = _get_pretty_string(list(self.resolvers.keys()))
        return f"{type(self).__name__}(scope={scope_keys}, resolvers={res_keys})"

    @property
    def has_resolvers(self) -> bool:
        """
        Return whether we have any extra scope.

        For example, DataFrames pass Their columns as resolvers during calls to
        ``DataFrame.eval()`` and ``DataFrame.query()``.

        Returns
        -------
        hr : bool
        """
        return bool(len(self.resolvers))

    def resolve(self, key: str, is_local: bool):
        """
        Resolve a variable name in a possibly local context.

        Parameters
        ----------
        key : str
            A variable name
        is_local : bool
            Flag indicating whether the variable is local or not (prefixed with
            the '@' symbol)

        Returns
        -------
        value : object
            The value of a particular variable
        """
        try:
            # only look for locals in outer scope
            if is_local:
                return self.scope[key]

            # not a local variable so check in resolvers if we have them
            if self.has_resolvers:
                return self.resolvers[key]

            # if we're here that means that we have no locals and we also have
            # no resolvers
            assert not is_local and not self.has_resolvers
            return self.scope[key]
        except KeyError:
            try:
                # last ditch effort we look in temporaries
                # these are created when parsing indexing expressions
                # e.g., df[df > 0]
                return self.temps[key]
            except KeyError as err:
                raise UndefinedVariableError(key, is_local) from err

    def swapkey(self, old_key: str, new_key: str, new_value=None) -> None:
        """
        Replace a variable name, with a potentially new value.

        Parameters
        ----------
        old_key : str
            Current variable name to replace
        new_key : str
            New variable name to replace `old_key` with
        new_value : object
            Value to be replaced along with the possible renaming
        """
        if self.has_resolvers:
            maps = self.resolvers.maps + self.scope.maps
        else:
            maps = self.scope.maps

        maps.append(self.temps)

        for mapping in maps:
            if old_key in mapping:
                mapping[new_key] = new_value
                return

    def _get_vars(self, stack, scopes: list[str]) -> None:
        """
        Get specifically scoped variables from a list of stack frames.

        Parameters
        ----------
        stack : list
            A list of stack frames as returned by ``inspect.stack()``
        scopes : sequence of strings
            A sequence containing valid stack frame attribute names that
            evaluate to a dictionary. For example, ('locals', 'globals')
        """
        variables = itertools.product(scopes, stack)
        for scope, (frame, _, _, _, _, _) in variables:
            try:
                d = getattr(frame, f"f_{scope}")
                self.scope = DeepChainMap(self.scope.new_child(d))
            finally:
                # won't remove it, but DECREF it
                # in Py3 this probably isn't necessary since frame won't be
                # scope after the loop
                del frame

    def _update(self, level: int) -> None:
        """
        Update the current scope by going back `level` levels.

        Parameters
        ----------
        level : int
        """
        sl = level + 1

        # add sl frames to the scope starting with the
        # most distant and overwriting with more current
        # makes sure that we can capture variable scope
        stack = inspect.stack()

        try:
            self._get_vars(stack[:sl], scopes=["locals"])
        finally:
            del stack[:], stack

    def add_tmp(self, value) -> str:
        """
        Add a temporary variable to the scope.

        Parameters
        ----------
        value : object
            An arbitrary object to be assigned to a temporary variable.

        Returns
        -------
        str
            The name of the temporary variable created.
        """
        name = f"{type(value).__name__}_{self.ntemps}_{_raw_hex_id(self)}"

        # add to inner most scope
        assert name not in self.temps
        self.temps[name] = value
        assert name in self.temps

        # only increment if the variable gets put in the scope
        return name

    @property
    def ntemps(self) -> int:
        """The number of temporary variables in this scope"""
        return len(self.temps)

    @property
    def full_scope(self) -> DeepChainMap:
        """
        Return the full scope for use with passing to engines transparently
        as a mapping.

        Returns
        -------
        vars : DeepChainMap
            All variables in this scope.
        """
        maps = [self.temps] + self.resolvers.maps + self.scope.maps
        return DeepChainMap(*maps)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\urllib3\_collections.py ===
from __future__ import absolute_import

try:
    from collections.abc import Mapping, MutableMapping
except ImportError:
    from collections import Mapping, MutableMapping
try:
    from threading import RLock
except ImportError:  # Platform-specific: No threads available

    class RLock:
        def __enter__(self):
            pass

        def __exit__(self, exc_type, exc_value, traceback):
            pass


from collections import OrderedDict

from .exceptions import InvalidHeader
from .packages import six
from .packages.six import iterkeys, itervalues

__all__ = ["RecentlyUsedContainer", "HTTPHeaderDict"]


_Null = object()


class RecentlyUsedContainer(MutableMapping):
    """
    Provides a thread-safe dict-like container which maintains up to
    ``maxsize`` keys while throwing away the least-recently-used keys beyond
    ``maxsize``.

    :param maxsize:
        Maximum number of recent elements to retain.

    :param dispose_func:
        Every time an item is evicted from the container,
        ``dispose_func(value)`` is called.  Callback which will get called
    """

    ContainerCls = OrderedDict

    def __init__(self, maxsize=10, dispose_func=None):
        self._maxsize = maxsize
        self.dispose_func = dispose_func

        self._container = self.ContainerCls()
        self.lock = RLock()

    def __getitem__(self, key):
        # Re-insert the item, moving it to the end of the eviction line.
        with self.lock:
            item = self._container.pop(key)
            self._container[key] = item
            return item

    def __setitem__(self, key, value):
        evicted_value = _Null
        with self.lock:
            # Possibly evict the existing value of 'key'
            evicted_value = self._container.get(key, _Null)
            self._container[key] = value

            # If we didn't evict an existing value, we might have to evict the
            # least recently used item from the beginning of the container.
            if len(self._container) > self._maxsize:
                _key, evicted_value = self._container.popitem(last=False)

        if self.dispose_func and evicted_value is not _Null:
            self.dispose_func(evicted_value)

    def __delitem__(self, key):
        with self.lock:
            value = self._container.pop(key)

        if self.dispose_func:
            self.dispose_func(value)

    def __len__(self):
        with self.lock:
            return len(self._container)

    def __iter__(self):
        raise NotImplementedError(
            "Iteration over this class is unlikely to be threadsafe."
        )

    def clear(self):
        with self.lock:
            # Copy pointers to all values, then wipe the mapping
            values = list(itervalues(self._container))
            self._container.clear()

        if self.dispose_func:
            for value in values:
                self.dispose_func(value)

    def keys(self):
        with self.lock:
            return list(iterkeys(self._container))


class HTTPHeaderDict(MutableMapping):
    """
    :param headers:
        An iterable of field-value pairs. Must not contain multiple field names
        when compared case-insensitively.

    :param kwargs:
        Additional field-value pairs to pass in to ``dict.update``.

    A ``dict`` like container for storing HTTP Headers.

    Field names are stored and compared case-insensitively in compliance with
    RFC 7230. Iteration provides the first case-sensitive key seen for each
    case-insensitive pair.

    Using ``__setitem__`` syntax overwrites fields that compare equal
    case-insensitively in order to maintain ``dict``'s api. For fields that
    compare equal, instead create a new ``HTTPHeaderDict`` and use ``.add``
    in a loop.

    If multiple fields that are equal case-insensitively are passed to the
    constructor or ``.update``, the behavior is undefined and some will be
    lost.

    >>> headers = HTTPHeaderDict()
    >>> headers.add('Set-Cookie', 'foo=bar')
    >>> headers.add('set-cookie', 'baz=quxx')
    >>> headers['content-length'] = '7'
    >>> headers['SET-cookie']
    'foo=bar, baz=quxx'
    >>> headers['Content-Length']
    '7'
    """

    def __init__(self, headers=None, **kwargs):
        super(HTTPHeaderDict, self).__init__()
        self._container = OrderedDict()
        if headers is not None:
            if isinstance(headers, HTTPHeaderDict):
                self._copy_from(headers)
            else:
                self.extend(headers)
        if kwargs:
            self.extend(kwargs)

    def __setitem__(self, key, val):
        self._container[key.lower()] = [key, val]
        return self._container[key.lower()]

    def __getitem__(self, key):
        val = self._container[key.lower()]
        return ", ".join(val[1:])

    def __delitem__(self, key):
        del self._container[key.lower()]

    def __contains__(self, key):
        return key.lower() in self._container

    def __eq__(self, other):
        if not isinstance(other, Mapping) and not hasattr(other, "keys"):
            return False
        if not isinstance(other, type(self)):
            other = type(self)(other)
        return dict((k.lower(), v) for k, v in self.itermerged()) == dict(
            (k.lower(), v) for k, v in other.itermerged()
        )

    def __ne__(self, other):
        return not self.__eq__(other)

    if six.PY2:  # Python 2
        iterkeys = MutableMapping.iterkeys
        itervalues = MutableMapping.itervalues

    __marker = object()

    def __len__(self):
        return len(self._container)

    def __iter__(self):
        # Only provide the originally cased names
        for vals in self._container.values():
            yield vals[0]

    def pop(self, key, default=__marker):
        """D.pop(k[,d]) -> v, remove specified key and return the corresponding value.
        If key is not found, d is returned if given, otherwise KeyError is raised.
        """
        # Using the MutableMapping function directly fails due to the private marker.
        # Using ordinary dict.pop would expose the internal structures.
        # So let's reinvent the wheel.
        try:
            value = self[key]
        except KeyError:
            if default is self.__marker:
                raise
            return default
        else:
            del self[key]
            return value

    def discard(self, key):
        try:
            del self[key]
        except KeyError:
            pass

    def add(self, key, val):
        """Adds a (name, value) pair, doesn't overwrite the value if it already
        exists.

        >>> headers = HTTPHeaderDict(foo='bar')
        >>> headers.add('Foo', 'baz')
        >>> headers['foo']
        'bar, baz'
        """
        key_lower = key.lower()
        new_vals = [key, val]
        # Keep the common case aka no item present as fast as possible
        vals = self._container.setdefault(key_lower, new_vals)
        if new_vals is not vals:
            vals.append(val)

    def extend(self, *args, **kwargs):
        """Generic import function for any type of header-like object.
        Adapted version of MutableMapping.update in order to insert items
        with self.add instead of self.__setitem__
        """
        if len(args) > 1:
            raise TypeError(
                "extend() takes at most 1 positional "
                "arguments ({0} given)".format(len(args))
            )
        other = args[0] if len(args) >= 1 else ()

        if isinstance(other, HTTPHeaderDict):
            for key, val in other.iteritems():
                self.add(key, val)
        elif isinstance(other, Mapping):
            for key in other:
                self.add(key, other[key])
        elif hasattr(other, "keys"):
            for key in other.keys():
                self.add(key, other[key])
        else:
            for key, value in other:
                self.add(key, value)

        for key, value in kwargs.items():
            self.add(key, value)

    def getlist(self, key, default=__marker):
        """Returns a list of all the values for the named field. Returns an
        empty list if the key doesn't exist."""
        try:
            vals = self._container[key.lower()]
        except KeyError:
            if default is self.__marker:
                return []
            return default
        else:
            return vals[1:]

    def _prepare_for_method_change(self):
        """
        Remove content-specific header fields before changing the request
        method to GET or HEAD according to RFC 9110, Section 15.4.
        """
        content_specific_headers = [
            "Content-Encoding",
            "Content-Language",
            "Content-Location",
            "Content-Type",
            "Content-Length",
            "Digest",
            "Last-Modified",
        ]
        for header in content_specific_headers:
            self.discard(header)
        return self

    # Backwards compatibility for httplib
    getheaders = getlist
    getallmatchingheaders = getlist
    iget = getlist

    # Backwards compatibility for http.cookiejar
    get_all = getlist

    def __repr__(self):
        return "%s(%s)" % (type(self).__name__, dict(self.itermerged()))

    def _copy_from(self, other):
        for key in other:
            val = other.getlist(key)
            if isinstance(val, list):
                # Don't need to convert tuples
                val = list(val)
            self._container[key.lower()] = [key] + val

    def copy(self):
        clone = type(self)()
        clone._copy_from(self)
        return clone

    def iteritems(self):
        """Iterate over all header lines, including duplicate ones."""
        for key in self:
            vals = self._container[key.lower()]
            for val in vals[1:]:
                yield vals[0], val

    def itermerged(self):
        """Iterate over all headers, merging duplicate ones together."""
        for key in self:
            val = self._container[key.lower()]
            yield val[0], ", ".join(val[1:])

    def items(self):
        return list(self.iteritems())

    @classmethod
    def from_httplib(cls, message):  # Python 2
        """Read headers from a Python 2 httplib message object."""
        # python2.7 does not expose a proper API for exporting multiheaders
        # efficiently. This function re-reads raw lines from the message
        # object and extracts the multiheaders properly.
        obs_fold_continued_leaders = (" ", "\t")
        headers = []

        for line in message.headers:
            if line.startswith(obs_fold_continued_leaders):
                if not headers:
                    # We received a header line that starts with OWS as described
                    # in RFC-7230 S3.2.4. This indicates a multiline header, but
                    # there exists no previous header to which we can attach it.
                    raise InvalidHeader(
                        "Header continuation with no previous header: %s" % line
                    )
                else:
                    key, value = headers[-1]
                    headers[-1] = (key, value + " " + line.strip())
                    continue

            key, value = line.split(":", 1)
            headers.append((key, value.strip()))

        return cls(headers)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\test_util.py ===
from __future__ import annotations

import sys
import types
from typing import TYPE_CHECKING, TypeVar

if TYPE_CHECKING:
    from collections.abc import AsyncGenerator, Coroutine, Generator
import pytest

import trio
from trio.testing import Matcher, RaisesGroup

from .. import _core
from .._core._tests.tutil import (
    create_asyncio_future_in_new_loop,
    ignore_coroutine_never_awaited_warnings,
)
from .._util import (
    ConflictDetector,
    MultipleExceptionError,
    NoPublicConstructor,
    coroutine_or_error,
    final,
    fixup_module_metadata,
    generic_function,
    is_main_thread,
    raise_single_exception_from_group,
)
from ..testing import wait_all_tasks_blocked

if sys.version_info < (3, 11):
    from exceptiongroup import BaseExceptionGroup, ExceptionGroup

if TYPE_CHECKING:
    from collections.abc import AsyncGenerator

T = TypeVar("T")


async def test_ConflictDetector() -> None:
    ul1 = ConflictDetector("ul1")
    ul2 = ConflictDetector("ul2")

    with ul1:
        with ul2:
            print("ok")

    with pytest.raises(_core.BusyResourceError, match="ul1"):
        with ul1:
            with ul1:
                pass  # pragma: no cover

    async def wait_with_ul1() -> None:
        with ul1:
            await wait_all_tasks_blocked()

    with RaisesGroup(Matcher(_core.BusyResourceError, "ul1")):
        async with _core.open_nursery() as nursery:
            nursery.start_soon(wait_with_ul1)
            nursery.start_soon(wait_with_ul1)


def test_module_metadata_is_fixed_up() -> None:
    import trio
    import trio.testing

    assert trio.Cancelled.__module__ == "trio"
    assert trio.open_nursery.__module__ == "trio"
    assert trio.abc.Stream.__module__ == "trio.abc"
    assert trio.lowlevel.wait_task_rescheduled.__module__ == "trio.lowlevel"
    assert trio.testing.trio_test.__module__ == "trio.testing"

    # Also check methods
    assert trio.lowlevel.ParkingLot.__init__.__module__ == "trio.lowlevel"
    assert trio.abc.Stream.send_all.__module__ == "trio.abc"

    # And names
    assert trio.Cancelled.__name__ == "Cancelled"
    assert trio.Cancelled.__qualname__ == "Cancelled"
    assert trio.abc.SendStream.send_all.__name__ == "send_all"
    assert trio.abc.SendStream.send_all.__qualname__ == "SendStream.send_all"
    assert trio.to_thread.__name__ == "trio.to_thread"
    assert trio.to_thread.run_sync.__name__ == "run_sync"
    assert trio.to_thread.run_sync.__qualname__ == "run_sync"


async def test_is_main_thread() -> None:
    assert is_main_thread()

    def not_main_thread() -> None:
        assert not is_main_thread()

    await trio.to_thread.run_sync(not_main_thread)


# @coroutine is deprecated since python 3.8, which is fine with us.
@pytest.mark.filterwarnings("ignore:.*@coroutine.*:DeprecationWarning")
def test_coroutine_or_error() -> None:
    class Deferred:
        "Just kidding"

    with ignore_coroutine_never_awaited_warnings():

        async def f() -> None:  # pragma: no cover
            pass

        with pytest.raises(TypeError) as excinfo:
            coroutine_or_error(f())  # type: ignore[arg-type, unused-coroutine]
        assert "expecting an async function" in str(excinfo.value)

        import asyncio

        if sys.version_info < (3, 11):

            @asyncio.coroutine
            def generator_based_coro() -> (
                Generator[Coroutine[None, None, None], None, None]
            ):  # pragma: no cover
                yield from asyncio.sleep(1)

            with pytest.raises(TypeError) as excinfo:
                coroutine_or_error(generator_based_coro())  # type: ignore[arg-type, unused-coroutine]
            assert "asyncio" in str(excinfo.value)

        with pytest.raises(TypeError) as excinfo:
            coroutine_or_error(create_asyncio_future_in_new_loop())  # type: ignore[arg-type, unused-coroutine]
        assert "asyncio" in str(excinfo.value)

        # does not raise arg-type error
        with pytest.raises(TypeError) as excinfo:
            coroutine_or_error(create_asyncio_future_in_new_loop)  # type: ignore[unused-coroutine]
        assert "asyncio" in str(excinfo.value)

        with pytest.raises(TypeError) as excinfo:
            coroutine_or_error(Deferred())  # type: ignore[arg-type, unused-coroutine]
        assert "twisted" in str(excinfo.value)

        with pytest.raises(TypeError) as excinfo:
            coroutine_or_error(lambda: Deferred())  # type: ignore[arg-type, unused-coroutine, return-value]
        assert "twisted" in str(excinfo.value)

        with pytest.raises(TypeError) as excinfo:
            coroutine_or_error(len, [[1, 2, 3]])  # type: ignore[arg-type, unused-coroutine]

        assert "appears to be synchronous" in str(excinfo.value)

        async def async_gen(
            _: object,
        ) -> AsyncGenerator[None, None]:  # pragma: no cover
            yield

        with pytest.raises(TypeError) as excinfo:
            coroutine_or_error(async_gen, [0])  # type: ignore[arg-type,unused-coroutine]
        msg = "expected an async function but got an async generator"
        assert msg in str(excinfo.value)

        # Make sure no references are kept around to keep anything alive
        del excinfo


def test_generic_function() -> None:
    @generic_function  # Decorated function contains "Any".
    def test_func(arg: T) -> T:  # type: ignore[misc]
        """Look, a docstring!"""
        return arg

    assert test_func is test_func[int] is test_func[int, str]
    assert test_func(42) == test_func[int](42) == 42
    assert test_func.__doc__ == "Look, a docstring!"
    assert test_func.__qualname__ == "test_generic_function.<locals>.test_func"  # type: ignore[attr-defined]
    assert test_func.__name__ == "test_func"  # type: ignore[attr-defined]
    assert test_func.__module__ == __name__


def test_final_decorator() -> None:
    """Test that subclassing a @final-annotated class is not allowed.

    This checks both runtime results, and verifies that type checkers detect
    the error statically through the type-ignore comment.
    """

    @final
    class FinalClass:
        pass

    with pytest.raises(TypeError):

        class SubClass(FinalClass):  # type: ignore[misc]
            pass


def test_no_public_constructor_metaclass() -> None:
    """The NoPublicConstructor metaclass prevents calling the constructor directly."""

    class SpecialClass(metaclass=NoPublicConstructor):
        def __init__(self, a: int, b: float) -> None:
            """Check arguments can be passed to __init__."""
            assert a == 8
            assert b == 3.15

    with pytest.raises(TypeError):
        SpecialClass(8, 3.15)

    # Private constructor should not raise, and passes args to __init__.
    assert isinstance(SpecialClass._create(8, b=3.15), SpecialClass)


def test_fixup_module_metadata() -> None:
    # Ignores modules not in the trio.X tree.
    non_trio_module = types.ModuleType("not_trio")
    non_trio_module.some_func = lambda: None  # type: ignore[attr-defined]
    non_trio_module.some_func.__name__ = "some_func"
    non_trio_module.some_func.__qualname__ = "some_func"

    fixup_module_metadata(non_trio_module.__name__, vars(non_trio_module))

    assert non_trio_module.some_func.__name__ == "some_func"
    assert non_trio_module.some_func.__qualname__ == "some_func"

    # Bulild up a fake module to test. Just use lambdas since all we care about is the names.
    mod = types.ModuleType("trio._somemodule_impl")
    mod.some_func = lambda: None  # type: ignore[attr-defined]
    mod.some_func.__name__ = "_something_else"
    mod.some_func.__qualname__ = "_something_else"

    # No __module__ means it's unchanged.
    mod.not_funclike = types.SimpleNamespace()  # type: ignore[attr-defined]
    mod.not_funclike.__name__ = "not_funclike"

    # Check __qualname__ being absent works.
    mod.only_has_name = types.SimpleNamespace()  # type: ignore[attr-defined]
    mod.only_has_name.__module__ = "trio._somemodule_impl"
    mod.only_has_name.__name__ = "only_name"

    # Underscored names are unchanged.
    mod._private = lambda: None  # type: ignore[attr-defined]
    mod._private.__module__ = "trio._somemodule_impl"
    mod._private.__name__ = mod._private.__qualname__ = "_private"

    # We recurse into classes.
    mod.SomeClass = type(  # type: ignore[attr-defined]
        "SomeClass",
        (),
        {
            "__init__": lambda self: None,
            "method": lambda self: None,
        },
    )
    # Reference loop is fine.
    mod.SomeClass.recursion = mod.SomeClass  # type: ignore[attr-defined]

    fixup_module_metadata("trio.somemodule", vars(mod))
    assert mod.some_func.__name__ == "some_func"
    assert mod.some_func.__module__ == "trio.somemodule"
    assert mod.some_func.__qualname__ == "some_func"

    assert mod.not_funclike.__name__ == "not_funclike"
    assert mod._private.__name__ == "_private"
    assert mod._private.__module__ == "trio._somemodule_impl"
    assert mod._private.__qualname__ == "_private"

    assert mod.only_has_name.__name__ == "only_has_name"
    assert mod.only_has_name.__module__ == "trio.somemodule"
    assert not hasattr(mod.only_has_name, "__qualname__")

    assert mod.SomeClass.method.__name__ == "method"  # type: ignore[attr-defined]
    assert mod.SomeClass.method.__module__ == "trio.somemodule"  # type: ignore[attr-defined]
    assert mod.SomeClass.method.__qualname__ == "SomeClass.method"  # type: ignore[attr-defined]
    # Make coverage happy.
    non_trio_module.some_func()
    mod.some_func()
    mod._private()
    mod.SomeClass().method()


async def test_raise_single_exception_from_group() -> None:
    excinfo: pytest.ExceptionInfo[BaseException]

    exc = ValueError("foo")
    cause = SyntaxError("cause")
    context = TypeError("context")
    exc.__cause__ = cause
    exc.__context__ = context
    cancelled = trio.Cancelled._create()

    with pytest.raises(ValueError, match="foo") as excinfo:
        raise_single_exception_from_group(ExceptionGroup("", [exc]))
    assert excinfo.value.__cause__ == cause
    assert excinfo.value.__context__ == context

    # only unwraps one layer of exceptiongroup
    inner_eg = ExceptionGroup("inner eg", [exc])
    inner_cause = SyntaxError("inner eg cause")
    inner_context = TypeError("inner eg context")
    inner_eg.__cause__ = inner_cause
    inner_eg.__context__ = inner_context
    with RaisesGroup(Matcher(ValueError, match="^foo$"), match="^inner eg$") as eginfo:
        raise_single_exception_from_group(ExceptionGroup("", [inner_eg]))
    assert eginfo.value.__cause__ == inner_cause
    assert eginfo.value.__context__ == inner_context

    with pytest.raises(ValueError, match="foo") as excinfo:
        raise_single_exception_from_group(
            BaseExceptionGroup("", [cancelled, cancelled, exc])
        )
    assert excinfo.value.__cause__ == cause
    assert excinfo.value.__context__ == context

    # multiple non-cancelled
    eg = ExceptionGroup("", [ValueError("foo"), ValueError("bar")])
    with pytest.raises(
        MultipleExceptionError,
        match=r"^Attempted to unwrap exceptiongroup with multiple non-cancelled exceptions. This is often caused by a bug in the caller.$",
    ) as excinfo:
        raise_single_exception_from_group(eg)
    assert excinfo.value.__cause__ is eg
    assert excinfo.value.__context__ is None

    # keyboardinterrupt overrides everything
    eg_ki = BaseExceptionGroup(
        "",
        [
            ValueError("foo"),
            ValueError("bar"),
            KeyboardInterrupt("this exc doesn't get reraised"),
        ],
    )
    with pytest.raises(KeyboardInterrupt, match=r"^$") as excinfo:
        raise_single_exception_from_group(eg_ki)
    assert excinfo.value.__cause__ is eg_ki
    assert excinfo.value.__context__ is None

    # and same for SystemExit
    systemexit_ki = BaseExceptionGroup(
        "",
        [
            ValueError("foo"),
            ValueError("bar"),
            SystemExit("this exc doesn't get reraised"),
        ],
    )
    with pytest.raises(SystemExit, match=r"^$") as excinfo:
        raise_single_exception_from_group(systemexit_ki)
    assert excinfo.value.__cause__ is systemexit_ki
    assert excinfo.value.__context__ is None

    # if we only got cancelled, first one is reraised
    with pytest.raises(trio.Cancelled, match=r"^Cancelled$") as excinfo:
        raise_single_exception_from_group(
            BaseExceptionGroup("", [cancelled, trio.Cancelled._create()])
        )
    assert excinfo.value is cancelled
    assert excinfo.value.__cause__ is None
    assert excinfo.value.__context__ is None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\concrete\expr_with_intlimits.py ===
from sympy.concrete.expr_with_limits import ExprWithLimits
from sympy.core.singleton import S
from sympy.core.relational import Eq

class ReorderError(NotImplementedError):
    """
    Exception raised when trying to reorder dependent limits.
    """
    def __init__(self, expr, msg):
        super().__init__(
            "%s could not be reordered: %s." % (expr, msg))

class ExprWithIntLimits(ExprWithLimits):
    """
    Superclass for Product and Sum.

    See Also
    ========

    sympy.concrete.expr_with_limits.ExprWithLimits
    sympy.concrete.products.Product
    sympy.concrete.summations.Sum
    """
    __slots__ = ()

    def change_index(self, var, trafo, newvar=None):
        r"""
        Change index of a Sum or Product.

        Perform a linear transformation `x \mapsto a x + b` on the index variable
        `x`. For `a` the only values allowed are `\pm 1`. A new variable to be used
        after the change of index can also be specified.

        Explanation
        ===========

        ``change_index(expr, var, trafo, newvar=None)`` where ``var`` specifies the
        index variable `x` to transform. The transformation ``trafo`` must be linear
        and given in terms of ``var``. If the optional argument ``newvar`` is
        provided then ``var`` gets replaced by ``newvar`` in the final expression.

        Examples
        ========

        >>> from sympy import Sum, Product, simplify
        >>> from sympy.abc import x, y, a, b, c, d, u, v, i, j, k, l

        >>> S = Sum(x, (x, a, b))
        >>> S.doit()
        -a**2/2 + a/2 + b**2/2 + b/2

        >>> Sn = S.change_index(x, x + 1, y)
        >>> Sn
        Sum(y - 1, (y, a + 1, b + 1))
        >>> Sn.doit()
        -a**2/2 + a/2 + b**2/2 + b/2

        >>> Sn = S.change_index(x, -x, y)
        >>> Sn
        Sum(-y, (y, -b, -a))
        >>> Sn.doit()
        -a**2/2 + a/2 + b**2/2 + b/2

        >>> Sn = S.change_index(x, x+u)
        >>> Sn
        Sum(-u + x, (x, a + u, b + u))
        >>> Sn.doit()
        -a**2/2 - a*u + a/2 + b**2/2 + b*u + b/2 - u*(-a + b + 1) + u
        >>> simplify(Sn.doit())
        -a**2/2 + a/2 + b**2/2 + b/2

        >>> Sn = S.change_index(x, -x - u, y)
        >>> Sn
        Sum(-u - y, (y, -b - u, -a - u))
        >>> Sn.doit()
        -a**2/2 - a*u + a/2 + b**2/2 + b*u + b/2 - u*(-a + b + 1) + u
        >>> simplify(Sn.doit())
        -a**2/2 + a/2 + b**2/2 + b/2

        >>> P = Product(i*j**2, (i, a, b), (j, c, d))
        >>> P
        Product(i*j**2, (i, a, b), (j, c, d))
        >>> P2 = P.change_index(i, i+3, k)
        >>> P2
        Product(j**2*(k - 3), (k, a + 3, b + 3), (j, c, d))
        >>> P3 = P2.change_index(j, -j, l)
        >>> P3
        Product(l**2*(k - 3), (k, a + 3, b + 3), (l, -d, -c))

        When dealing with symbols only, we can make a
        general linear transformation:

        >>> Sn = S.change_index(x, u*x+v, y)
        >>> Sn
        Sum((-v + y)/u, (y, b*u + v, a*u + v))
        >>> Sn.doit()
        -v*(a*u - b*u + 1)/u + (a**2*u**2/2 + a*u*v + a*u/2 - b**2*u**2/2 - b*u*v + b*u/2 + v)/u
        >>> simplify(Sn.doit())
        a**2*u/2 + a/2 - b**2*u/2 + b/2

        However, the last result can be inconsistent with usual
        summation where the index increment is always 1. This is
        obvious as we get back the original value only for ``u``
        equal +1 or -1.

        See Also
        ========

        sympy.concrete.expr_with_intlimits.ExprWithIntLimits.index,
        reorder_limit,
        sympy.concrete.expr_with_intlimits.ExprWithIntLimits.reorder,
        sympy.concrete.summations.Sum.reverse_order,
        sympy.concrete.products.Product.reverse_order
        """
        if newvar is None:
            newvar = var

        limits = []
        for limit in self.limits:
            if limit[0] == var:
                p = trafo.as_poly(var)
                if p.degree() != 1:
                    raise ValueError("Index transformation is not linear")
                alpha = p.coeff_monomial(var)
                beta = p.coeff_monomial(S.One)
                if alpha.is_number:
                    if alpha == S.One:
                        limits.append((newvar, alpha*limit[1] + beta, alpha*limit[2] + beta))
                    elif alpha == S.NegativeOne:
                        limits.append((newvar, alpha*limit[2] + beta, alpha*limit[1] + beta))
                    else:
                        raise ValueError("Linear transformation results in non-linear summation stepsize")
                else:
                    # Note that the case of alpha being symbolic can give issues if alpha < 0.
                    limits.append((newvar, alpha*limit[2] + beta, alpha*limit[1] + beta))
            else:
                limits.append(limit)

        function = self.function.subs(var, (var - beta)/alpha)
        function = function.subs(var, newvar)

        return self.func(function, *limits)


    def index(expr, x):
        """
        Return the index of a dummy variable in the list of limits.

        Explanation
        ===========

        ``index(expr, x)``  returns the index of the dummy variable ``x`` in the
        limits of ``expr``. Note that we start counting with 0 at the inner-most
        limits tuple.

        Examples
        ========

        >>> from sympy.abc import x, y, a, b, c, d
        >>> from sympy import Sum, Product
        >>> Sum(x*y, (x, a, b), (y, c, d)).index(x)
        0
        >>> Sum(x*y, (x, a, b), (y, c, d)).index(y)
        1
        >>> Product(x*y, (x, a, b), (y, c, d)).index(x)
        0
        >>> Product(x*y, (x, a, b), (y, c, d)).index(y)
        1

        See Also
        ========

        reorder_limit, reorder, sympy.concrete.summations.Sum.reverse_order,
        sympy.concrete.products.Product.reverse_order
        """
        variables = [limit[0] for limit in expr.limits]

        if variables.count(x) != 1:
            raise ValueError(expr, "Number of instances of variable not equal to one")
        else:
            return variables.index(x)

    def reorder(expr, *arg):
        """
        Reorder limits in a expression containing a Sum or a Product.

        Explanation
        ===========

        ``expr.reorder(*arg)`` reorders the limits in the expression ``expr``
        according to the list of tuples given by ``arg``. These tuples can
        contain numerical indices or index variable names or involve both.

        Examples
        ========

        >>> from sympy import Sum, Product
        >>> from sympy.abc import x, y, z, a, b, c, d, e, f

        >>> Sum(x*y, (x, a, b), (y, c, d)).reorder((x, y))
        Sum(x*y, (y, c, d), (x, a, b))

        >>> Sum(x*y*z, (x, a, b), (y, c, d), (z, e, f)).reorder((x, y), (x, z), (y, z))
        Sum(x*y*z, (z, e, f), (y, c, d), (x, a, b))

        >>> P = Product(x*y*z, (x, a, b), (y, c, d), (z, e, f))
        >>> P.reorder((x, y), (x, z), (y, z))
        Product(x*y*z, (z, e, f), (y, c, d), (x, a, b))

        We can also select the index variables by counting them, starting
        with the inner-most one:

        >>> Sum(x**2, (x, a, b), (x, c, d)).reorder((0, 1))
        Sum(x**2, (x, c, d), (x, a, b))

        And of course we can mix both schemes:

        >>> Sum(x*y, (x, a, b), (y, c, d)).reorder((y, x))
        Sum(x*y, (y, c, d), (x, a, b))
        >>> Sum(x*y, (x, a, b), (y, c, d)).reorder((y, 0))
        Sum(x*y, (y, c, d), (x, a, b))

        See Also
        ========

        reorder_limit, index, sympy.concrete.summations.Sum.reverse_order,
        sympy.concrete.products.Product.reverse_order
        """
        new_expr = expr

        for r in arg:
            if len(r) != 2:
                raise ValueError(r, "Invalid number of arguments")

            index1 = r[0]
            index2 = r[1]

            if not isinstance(r[0], int):
                index1 = expr.index(r[0])
            if not isinstance(r[1], int):
                index2 = expr.index(r[1])

            new_expr = new_expr.reorder_limit(index1, index2)

        return new_expr


    def reorder_limit(expr, x, y):
        """
        Interchange two limit tuples of a Sum or Product expression.

        Explanation
        ===========

        ``expr.reorder_limit(x, y)`` interchanges two limit tuples. The
        arguments ``x`` and ``y`` are integers corresponding to the index
        variables of the two limits which are to be interchanged. The
        expression ``expr`` has to be either a Sum or a Product.

        Examples
        ========

        >>> from sympy.abc import x, y, z, a, b, c, d, e, f
        >>> from sympy import Sum, Product

        >>> Sum(x*y*z, (x, a, b), (y, c, d), (z, e, f)).reorder_limit(0, 2)
        Sum(x*y*z, (z, e, f), (y, c, d), (x, a, b))
        >>> Sum(x**2, (x, a, b), (x, c, d)).reorder_limit(1, 0)
        Sum(x**2, (x, c, d), (x, a, b))

        >>> Product(x*y*z, (x, a, b), (y, c, d), (z, e, f)).reorder_limit(0, 2)
        Product(x*y*z, (z, e, f), (y, c, d), (x, a, b))

        See Also
        ========

        index, reorder, sympy.concrete.summations.Sum.reverse_order,
        sympy.concrete.products.Product.reverse_order
        """
        var = {limit[0] for limit in expr.limits}
        limit_x = expr.limits[x]
        limit_y = expr.limits[y]

        if (len(set(limit_x[1].free_symbols).intersection(var)) == 0 and
            len(set(limit_x[2].free_symbols).intersection(var)) == 0 and
            len(set(limit_y[1].free_symbols).intersection(var)) == 0 and
            len(set(limit_y[2].free_symbols).intersection(var)) == 0):

            limits = []
            for i, limit in enumerate(expr.limits):
                if i == x:
                    limits.append(limit_y)
                elif i == y:
                    limits.append(limit_x)
                else:
                    limits.append(limit)

            return type(expr)(expr.function, *limits)
        else:
            raise ReorderError(expr, "could not interchange the two limits specified")

    @property
    def has_empty_sequence(self):
        """
        Returns True if the Sum or Product is computed for an empty sequence.

        Examples
        ========

        >>> from sympy import Sum, Product, Symbol
        >>> m = Symbol('m')
        >>> Sum(m, (m, 1, 0)).has_empty_sequence
        True

        >>> Sum(m, (m, 1, 1)).has_empty_sequence
        False

        >>> M = Symbol('M', integer=True, positive=True)
        >>> Product(m, (m, 1, M)).has_empty_sequence
        False

        >>> Product(m, (m, 2, M)).has_empty_sequence

        >>> Product(m, (m, M + 1, M)).has_empty_sequence
        True

        >>> N = Symbol('N', integer=True, positive=True)
        >>> Sum(m, (m, N, M)).has_empty_sequence

        >>> N = Symbol('N', integer=True, negative=True)
        >>> Sum(m, (m, N, M)).has_empty_sequence
        False

        See Also
        ========

        has_reversed_limits
        has_finite_limits

        """
        ret_None = False
        for lim in self.limits:
            dif = lim[1] - lim[2]
            eq = Eq(dif, 1)
            if eq == True:
                return True
            elif eq == False:
                continue
            else:
                ret_None = True

        if ret_None:
            return None
        return False

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\packaging\_parser.py ===
"""Handwritten parser of dependency specifiers.

The docstring for each __parse_* function contains EBNF-inspired grammar representing
the implementation.
"""

from __future__ import annotations

import ast
from typing import NamedTuple, Sequence, Tuple, Union

from ._tokenizer import DEFAULT_RULES, Tokenizer


class Node:
    def __init__(self, value: str) -> None:
        self.value = value

    def __str__(self) -> str:
        return self.value

    def __repr__(self) -> str:
        return f"<{self.__class__.__name__}('{self}')>"

    def serialize(self) -> str:
        raise NotImplementedError


class Variable(Node):
    def serialize(self) -> str:
        return str(self)


class Value(Node):
    def serialize(self) -> str:
        return f'"{self}"'


class Op(Node):
    def serialize(self) -> str:
        return str(self)


MarkerVar = Union[Variable, Value]
MarkerItem = Tuple[MarkerVar, Op, MarkerVar]
MarkerAtom = Union[MarkerItem, Sequence["MarkerAtom"]]
MarkerList = Sequence[Union["MarkerList", MarkerAtom, str]]


class ParsedRequirement(NamedTuple):
    name: str
    url: str
    extras: list[str]
    specifier: str
    marker: MarkerList | None


# --------------------------------------------------------------------------------------
# Recursive descent parser for dependency specifier
# --------------------------------------------------------------------------------------
def parse_requirement(source: str) -> ParsedRequirement:
    return _parse_requirement(Tokenizer(source, rules=DEFAULT_RULES))


def _parse_requirement(tokenizer: Tokenizer) -> ParsedRequirement:
    """
    requirement = WS? IDENTIFIER WS? extras WS? requirement_details
    """
    tokenizer.consume("WS")

    name_token = tokenizer.expect(
        "IDENTIFIER", expected="package name at the start of dependency specifier"
    )
    name = name_token.text
    tokenizer.consume("WS")

    extras = _parse_extras(tokenizer)
    tokenizer.consume("WS")

    url, specifier, marker = _parse_requirement_details(tokenizer)
    tokenizer.expect("END", expected="end of dependency specifier")

    return ParsedRequirement(name, url, extras, specifier, marker)


def _parse_requirement_details(
    tokenizer: Tokenizer,
) -> tuple[str, str, MarkerList | None]:
    """
    requirement_details = AT URL (WS requirement_marker?)?
                        | specifier WS? (requirement_marker)?
    """

    specifier = ""
    url = ""
    marker = None

    if tokenizer.check("AT"):
        tokenizer.read()
        tokenizer.consume("WS")

        url_start = tokenizer.position
        url = tokenizer.expect("URL", expected="URL after @").text
        if tokenizer.check("END", peek=True):
            return (url, specifier, marker)

        tokenizer.expect("WS", expected="whitespace after URL")

        # The input might end after whitespace.
        if tokenizer.check("END", peek=True):
            return (url, specifier, marker)

        marker = _parse_requirement_marker(
            tokenizer, span_start=url_start, after="URL and whitespace"
        )
    else:
        specifier_start = tokenizer.position
        specifier = _parse_specifier(tokenizer)
        tokenizer.consume("WS")

        if tokenizer.check("END", peek=True):
            return (url, specifier, marker)

        marker = _parse_requirement_marker(
            tokenizer,
            span_start=specifier_start,
            after=(
                "version specifier"
                if specifier
                else "name and no valid version specifier"
            ),
        )

    return (url, specifier, marker)


def _parse_requirement_marker(
    tokenizer: Tokenizer, *, span_start: int, after: str
) -> MarkerList:
    """
    requirement_marker = SEMICOLON marker WS?
    """

    if not tokenizer.check("SEMICOLON"):
        tokenizer.raise_syntax_error(
            f"Expected end or semicolon (after {after})",
            span_start=span_start,
        )
    tokenizer.read()

    marker = _parse_marker(tokenizer)
    tokenizer.consume("WS")

    return marker


def _parse_extras(tokenizer: Tokenizer) -> list[str]:
    """
    extras = (LEFT_BRACKET wsp* extras_list? wsp* RIGHT_BRACKET)?
    """
    if not tokenizer.check("LEFT_BRACKET", peek=True):
        return []

    with tokenizer.enclosing_tokens(
        "LEFT_BRACKET",
        "RIGHT_BRACKET",
        around="extras",
    ):
        tokenizer.consume("WS")
        extras = _parse_extras_list(tokenizer)
        tokenizer.consume("WS")

    return extras


def _parse_extras_list(tokenizer: Tokenizer) -> list[str]:
    """
    extras_list = identifier (wsp* ',' wsp* identifier)*
    """
    extras: list[str] = []

    if not tokenizer.check("IDENTIFIER"):
        return extras

    extras.append(tokenizer.read().text)

    while True:
        tokenizer.consume("WS")
        if tokenizer.check("IDENTIFIER", peek=True):
            tokenizer.raise_syntax_error("Expected comma between extra names")
        elif not tokenizer.check("COMMA"):
            break

        tokenizer.read()
        tokenizer.consume("WS")

        extra_token = tokenizer.expect("IDENTIFIER", expected="extra name after comma")
        extras.append(extra_token.text)

    return extras


def _parse_specifier(tokenizer: Tokenizer) -> str:
    """
    specifier = LEFT_PARENTHESIS WS? version_many WS? RIGHT_PARENTHESIS
              | WS? version_many WS?
    """
    with tokenizer.enclosing_tokens(
        "LEFT_PARENTHESIS",
        "RIGHT_PARENTHESIS",
        around="version specifier",
    ):
        tokenizer.consume("WS")
        parsed_specifiers = _parse_version_many(tokenizer)
        tokenizer.consume("WS")

    return parsed_specifiers


def _parse_version_many(tokenizer: Tokenizer) -> str:
    """
    version_many = (SPECIFIER (WS? COMMA WS? SPECIFIER)*)?
    """
    parsed_specifiers = ""
    while tokenizer.check("SPECIFIER"):
        span_start = tokenizer.position
        parsed_specifiers += tokenizer.read().text
        if tokenizer.check("VERSION_PREFIX_TRAIL", peek=True):
            tokenizer.raise_syntax_error(
                ".* suffix can only be used with `==` or `!=` operators",
                span_start=span_start,
                span_end=tokenizer.position + 1,
            )
        if tokenizer.check("VERSION_LOCAL_LABEL_TRAIL", peek=True):
            tokenizer.raise_syntax_error(
                "Local version label can only be used with `==` or `!=` operators",
                span_start=span_start,
                span_end=tokenizer.position,
            )
        tokenizer.consume("WS")
        if not tokenizer.check("COMMA"):
            break
        parsed_specifiers += tokenizer.read().text
        tokenizer.consume("WS")

    return parsed_specifiers


# --------------------------------------------------------------------------------------
# Recursive descent parser for marker expression
# --------------------------------------------------------------------------------------
def parse_marker(source: str) -> MarkerList:
    return _parse_full_marker(Tokenizer(source, rules=DEFAULT_RULES))


def _parse_full_marker(tokenizer: Tokenizer) -> MarkerList:
    retval = _parse_marker(tokenizer)
    tokenizer.expect("END", expected="end of marker expression")
    return retval


def _parse_marker(tokenizer: Tokenizer) -> MarkerList:
    """
    marker = marker_atom (BOOLOP marker_atom)+
    """
    expression = [_parse_marker_atom(tokenizer)]
    while tokenizer.check("BOOLOP"):
        token = tokenizer.read()
        expr_right = _parse_marker_atom(tokenizer)
        expression.extend((token.text, expr_right))
    return expression


def _parse_marker_atom(tokenizer: Tokenizer) -> MarkerAtom:
    """
    marker_atom = WS? LEFT_PARENTHESIS WS? marker WS? RIGHT_PARENTHESIS WS?
                | WS? marker_item WS?
    """

    tokenizer.consume("WS")
    if tokenizer.check("LEFT_PARENTHESIS", peek=True):
        with tokenizer.enclosing_tokens(
            "LEFT_PARENTHESIS",
            "RIGHT_PARENTHESIS",
            around="marker expression",
        ):
            tokenizer.consume("WS")
            marker: MarkerAtom = _parse_marker(tokenizer)
            tokenizer.consume("WS")
    else:
        marker = _parse_marker_item(tokenizer)
    tokenizer.consume("WS")
    return marker


def _parse_marker_item(tokenizer: Tokenizer) -> MarkerItem:
    """
    marker_item = WS? marker_var WS? marker_op WS? marker_var WS?
    """
    tokenizer.consume("WS")
    marker_var_left = _parse_marker_var(tokenizer)
    tokenizer.consume("WS")
    marker_op = _parse_marker_op(tokenizer)
    tokenizer.consume("WS")
    marker_var_right = _parse_marker_var(tokenizer)
    tokenizer.consume("WS")
    return (marker_var_left, marker_op, marker_var_right)


def _parse_marker_var(tokenizer: Tokenizer) -> MarkerVar:
    """
    marker_var = VARIABLE | QUOTED_STRING
    """
    if tokenizer.check("VARIABLE"):
        return process_env_var(tokenizer.read().text.replace(".", "_"))
    elif tokenizer.check("QUOTED_STRING"):
        return process_python_str(tokenizer.read().text)
    else:
        tokenizer.raise_syntax_error(
            message="Expected a marker variable or quoted string"
        )


def process_env_var(env_var: str) -> Variable:
    if env_var in ("platform_python_implementation", "python_implementation"):
        return Variable("platform_python_implementation")
    else:
        return Variable(env_var)


def process_python_str(python_str: str) -> Value:
    value = ast.literal_eval(python_str)
    return Value(str(value))


def _parse_marker_op(tokenizer: Tokenizer) -> Op:
    """
    marker_op = IN | NOT IN | OP
    """
    if tokenizer.check("IN"):
        tokenizer.read()
        return Op("in")
    elif tokenizer.check("NOT"):
        tokenizer.read()
        tokenizer.expect("WS", expected="whitespace after 'not'")
        tokenizer.expect("IN", expected="'in' after 'not'")
        return Op("not in")
    elif tokenizer.check("OP"):
        return Op(tokenizer.read().text)
    else:
        return tokenizer.raise_syntax_error(
            "Expected marker operator, one of <=, <, !=, ==, >=, >, ~=, ===, in, not in"
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\packaging\_parser.py ===
"""Handwritten parser of dependency specifiers.

The docstring for each __parse_* function contains EBNF-inspired grammar representing
the implementation.
"""

from __future__ import annotations

import ast
from typing import NamedTuple, Sequence, Tuple, Union

from ._tokenizer import DEFAULT_RULES, Tokenizer


class Node:
    def __init__(self, value: str) -> None:
        self.value = value

    def __str__(self) -> str:
        return self.value

    def __repr__(self) -> str:
        return f"<{self.__class__.__name__}('{self}')>"

    def serialize(self) -> str:
        raise NotImplementedError


class Variable(Node):
    def serialize(self) -> str:
        return str(self)


class Value(Node):
    def serialize(self) -> str:
        return f'"{self}"'


class Op(Node):
    def serialize(self) -> str:
        return str(self)


MarkerVar = Union[Variable, Value]
MarkerItem = Tuple[MarkerVar, Op, MarkerVar]
MarkerAtom = Union[MarkerItem, Sequence["MarkerAtom"]]
MarkerList = Sequence[Union["MarkerList", MarkerAtom, str]]


class ParsedRequirement(NamedTuple):
    name: str
    url: str
    extras: list[str]
    specifier: str
    marker: MarkerList | None


# --------------------------------------------------------------------------------------
# Recursive descent parser for dependency specifier
# --------------------------------------------------------------------------------------
def parse_requirement(source: str) -> ParsedRequirement:
    return _parse_requirement(Tokenizer(source, rules=DEFAULT_RULES))


def _parse_requirement(tokenizer: Tokenizer) -> ParsedRequirement:
    """
    requirement = WS? IDENTIFIER WS? extras WS? requirement_details
    """
    tokenizer.consume("WS")

    name_token = tokenizer.expect(
        "IDENTIFIER", expected="package name at the start of dependency specifier"
    )
    name = name_token.text
    tokenizer.consume("WS")

    extras = _parse_extras(tokenizer)
    tokenizer.consume("WS")

    url, specifier, marker = _parse_requirement_details(tokenizer)
    tokenizer.expect("END", expected="end of dependency specifier")

    return ParsedRequirement(name, url, extras, specifier, marker)


def _parse_requirement_details(
    tokenizer: Tokenizer,
) -> tuple[str, str, MarkerList | None]:
    """
    requirement_details = AT URL (WS requirement_marker?)?
                        | specifier WS? (requirement_marker)?
    """

    specifier = ""
    url = ""
    marker = None

    if tokenizer.check("AT"):
        tokenizer.read()
        tokenizer.consume("WS")

        url_start = tokenizer.position
        url = tokenizer.expect("URL", expected="URL after @").text
        if tokenizer.check("END", peek=True):
            return (url, specifier, marker)

        tokenizer.expect("WS", expected="whitespace after URL")

        # The input might end after whitespace.
        if tokenizer.check("END", peek=True):
            return (url, specifier, marker)

        marker = _parse_requirement_marker(
            tokenizer, span_start=url_start, after="URL and whitespace"
        )
    else:
        specifier_start = tokenizer.position
        specifier = _parse_specifier(tokenizer)
        tokenizer.consume("WS")

        if tokenizer.check("END", peek=True):
            return (url, specifier, marker)

        marker = _parse_requirement_marker(
            tokenizer,
            span_start=specifier_start,
            after=(
                "version specifier"
                if specifier
                else "name and no valid version specifier"
            ),
        )

    return (url, specifier, marker)


def _parse_requirement_marker(
    tokenizer: Tokenizer, *, span_start: int, after: str
) -> MarkerList:
    """
    requirement_marker = SEMICOLON marker WS?
    """

    if not tokenizer.check("SEMICOLON"):
        tokenizer.raise_syntax_error(
            f"Expected end or semicolon (after {after})",
            span_start=span_start,
        )
    tokenizer.read()

    marker = _parse_marker(tokenizer)
    tokenizer.consume("WS")

    return marker


def _parse_extras(tokenizer: Tokenizer) -> list[str]:
    """
    extras = (LEFT_BRACKET wsp* extras_list? wsp* RIGHT_BRACKET)?
    """
    if not tokenizer.check("LEFT_BRACKET", peek=True):
        return []

    with tokenizer.enclosing_tokens(
        "LEFT_BRACKET",
        "RIGHT_BRACKET",
        around="extras",
    ):
        tokenizer.consume("WS")
        extras = _parse_extras_list(tokenizer)
        tokenizer.consume("WS")

    return extras


def _parse_extras_list(tokenizer: Tokenizer) -> list[str]:
    """
    extras_list = identifier (wsp* ',' wsp* identifier)*
    """
    extras: list[str] = []

    if not tokenizer.check("IDENTIFIER"):
        return extras

    extras.append(tokenizer.read().text)

    while True:
        tokenizer.consume("WS")
        if tokenizer.check("IDENTIFIER", peek=True):
            tokenizer.raise_syntax_error("Expected comma between extra names")
        elif not tokenizer.check("COMMA"):
            break

        tokenizer.read()
        tokenizer.consume("WS")

        extra_token = tokenizer.expect("IDENTIFIER", expected="extra name after comma")
        extras.append(extra_token.text)

    return extras


def _parse_specifier(tokenizer: Tokenizer) -> str:
    """
    specifier = LEFT_PARENTHESIS WS? version_many WS? RIGHT_PARENTHESIS
              | WS? version_many WS?
    """
    with tokenizer.enclosing_tokens(
        "LEFT_PARENTHESIS",
        "RIGHT_PARENTHESIS",
        around="version specifier",
    ):
        tokenizer.consume("WS")
        parsed_specifiers = _parse_version_many(tokenizer)
        tokenizer.consume("WS")

    return parsed_specifiers


def _parse_version_many(tokenizer: Tokenizer) -> str:
    """
    version_many = (SPECIFIER (WS? COMMA WS? SPECIFIER)*)?
    """
    parsed_specifiers = ""
    while tokenizer.check("SPECIFIER"):
        span_start = tokenizer.position
        parsed_specifiers += tokenizer.read().text
        if tokenizer.check("VERSION_PREFIX_TRAIL", peek=True):
            tokenizer.raise_syntax_error(
                ".* suffix can only be used with `==` or `!=` operators",
                span_start=span_start,
                span_end=tokenizer.position + 1,
            )
        if tokenizer.check("VERSION_LOCAL_LABEL_TRAIL", peek=True):
            tokenizer.raise_syntax_error(
                "Local version label can only be used with `==` or `!=` operators",
                span_start=span_start,
                span_end=tokenizer.position,
            )
        tokenizer.consume("WS")
        if not tokenizer.check("COMMA"):
            break
        parsed_specifiers += tokenizer.read().text
        tokenizer.consume("WS")

    return parsed_specifiers


# --------------------------------------------------------------------------------------
# Recursive descent parser for marker expression
# --------------------------------------------------------------------------------------
def parse_marker(source: str) -> MarkerList:
    return _parse_full_marker(Tokenizer(source, rules=DEFAULT_RULES))


def _parse_full_marker(tokenizer: Tokenizer) -> MarkerList:
    retval = _parse_marker(tokenizer)
    tokenizer.expect("END", expected="end of marker expression")
    return retval


def _parse_marker(tokenizer: Tokenizer) -> MarkerList:
    """
    marker = marker_atom (BOOLOP marker_atom)+
    """
    expression = [_parse_marker_atom(tokenizer)]
    while tokenizer.check("BOOLOP"):
        token = tokenizer.read()
        expr_right = _parse_marker_atom(tokenizer)
        expression.extend((token.text, expr_right))
    return expression


def _parse_marker_atom(tokenizer: Tokenizer) -> MarkerAtom:
    """
    marker_atom = WS? LEFT_PARENTHESIS WS? marker WS? RIGHT_PARENTHESIS WS?
                | WS? marker_item WS?
    """

    tokenizer.consume("WS")
    if tokenizer.check("LEFT_PARENTHESIS", peek=True):
        with tokenizer.enclosing_tokens(
            "LEFT_PARENTHESIS",
            "RIGHT_PARENTHESIS",
            around="marker expression",
        ):
            tokenizer.consume("WS")
            marker: MarkerAtom = _parse_marker(tokenizer)
            tokenizer.consume("WS")
    else:
        marker = _parse_marker_item(tokenizer)
    tokenizer.consume("WS")
    return marker


def _parse_marker_item(tokenizer: Tokenizer) -> MarkerItem:
    """
    marker_item = WS? marker_var WS? marker_op WS? marker_var WS?
    """
    tokenizer.consume("WS")
    marker_var_left = _parse_marker_var(tokenizer)
    tokenizer.consume("WS")
    marker_op = _parse_marker_op(tokenizer)
    tokenizer.consume("WS")
    marker_var_right = _parse_marker_var(tokenizer)
    tokenizer.consume("WS")
    return (marker_var_left, marker_op, marker_var_right)


def _parse_marker_var(tokenizer: Tokenizer) -> MarkerVar:
    """
    marker_var = VARIABLE | QUOTED_STRING
    """
    if tokenizer.check("VARIABLE"):
        return process_env_var(tokenizer.read().text.replace(".", "_"))
    elif tokenizer.check("QUOTED_STRING"):
        return process_python_str(tokenizer.read().text)
    else:
        tokenizer.raise_syntax_error(
            message="Expected a marker variable or quoted string"
        )


def process_env_var(env_var: str) -> Variable:
    if env_var in ("platform_python_implementation", "python_implementation"):
        return Variable("platform_python_implementation")
    else:
        return Variable(env_var)


def process_python_str(python_str: str) -> Value:
    value = ast.literal_eval(python_str)
    return Value(str(value))


def _parse_marker_op(tokenizer: Tokenizer) -> Op:
    """
    marker_op = IN | NOT IN | OP
    """
    if tokenizer.check("IN"):
        tokenizer.read()
        return Op("in")
    elif tokenizer.check("NOT"):
        tokenizer.read()
        tokenizer.expect("WS", expected="whitespace after 'not'")
        tokenizer.expect("IN", expected="'in' after 'not'")
        return Op("not in")
    elif tokenizer.check("OP"):
        return Op(tokenizer.read().text)
    else:
        return tokenizer.raise_syntax_error(
            "Expected marker operator, one of <=, <, !=, ==, >=, >, ~=, ===, in, not in"
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\printing\mathematica.py ===
"""
Mathematica code printer
"""

from __future__ import annotations
from typing import Any

from sympy.core import Basic, Expr, Float
from sympy.core.sorting import default_sort_key

from sympy.printing.codeprinter import CodePrinter
from sympy.printing.precedence import precedence

# Used in MCodePrinter._print_Function(self)
known_functions = {
    "exp": [(lambda x: True, "Exp")],
    "log": [(lambda x: True, "Log")],
    "sin": [(lambda x: True, "Sin")],
    "cos": [(lambda x: True, "Cos")],
    "tan": [(lambda x: True, "Tan")],
    "cot": [(lambda x: True, "Cot")],
    "sec": [(lambda x: True, "Sec")],
    "csc": [(lambda x: True, "Csc")],
    "asin": [(lambda x: True, "ArcSin")],
    "acos": [(lambda x: True, "ArcCos")],
    "atan": [(lambda x: True, "ArcTan")],
    "acot": [(lambda x: True, "ArcCot")],
    "asec": [(lambda x: True, "ArcSec")],
    "acsc": [(lambda x: True, "ArcCsc")],
    "sinh": [(lambda x: True, "Sinh")],
    "cosh": [(lambda x: True, "Cosh")],
    "tanh": [(lambda x: True, "Tanh")],
    "coth": [(lambda x: True, "Coth")],
    "sech": [(lambda x: True, "Sech")],
    "csch": [(lambda x: True, "Csch")],
    "asinh": [(lambda x: True, "ArcSinh")],
    "acosh": [(lambda x: True, "ArcCosh")],
    "atanh": [(lambda x: True, "ArcTanh")],
    "acoth": [(lambda x: True, "ArcCoth")],
    "asech": [(lambda x: True, "ArcSech")],
    "acsch": [(lambda x: True, "ArcCsch")],
    "sinc": [(lambda x: True, "Sinc")],
    "conjugate": [(lambda x: True, "Conjugate")],
    "Max": [(lambda *x: True, "Max")],
    "Min": [(lambda *x: True, "Min")],
    "erf": [(lambda x: True, "Erf")],
    "erf2": [(lambda *x: True, "Erf")],
    "erfc": [(lambda x: True, "Erfc")],
    "erfi": [(lambda x: True, "Erfi")],
    "erfinv": [(lambda x: True, "InverseErf")],
    "erfcinv": [(lambda x: True, "InverseErfc")],
    "erf2inv": [(lambda *x: True, "InverseErf")],
    "expint": [(lambda *x: True, "ExpIntegralE")],
    "Ei": [(lambda x: True, "ExpIntegralEi")],
    "fresnelc": [(lambda x: True, "FresnelC")],
    "fresnels": [(lambda x: True, "FresnelS")],
    "gamma": [(lambda x: True, "Gamma")],
    "uppergamma": [(lambda *x: True, "Gamma")],
    "polygamma": [(lambda *x: True, "PolyGamma")],
    "loggamma": [(lambda x: True, "LogGamma")],
    "beta": [(lambda *x: True, "Beta")],
    "Ci": [(lambda x: True, "CosIntegral")],
    "Si": [(lambda x: True, "SinIntegral")],
    "Chi": [(lambda x: True, "CoshIntegral")],
    "Shi": [(lambda x: True, "SinhIntegral")],
    "li": [(lambda x: True, "LogIntegral")],
    "factorial": [(lambda x: True, "Factorial")],
    "factorial2": [(lambda x: True, "Factorial2")],
    "subfactorial": [(lambda x: True, "Subfactorial")],
    "catalan": [(lambda x: True, "CatalanNumber")],
    "harmonic": [(lambda *x: True, "HarmonicNumber")],
    "lucas": [(lambda x: True, "LucasL")],
    "RisingFactorial": [(lambda *x: True, "Pochhammer")],
    "FallingFactorial": [(lambda *x: True, "FactorialPower")],
    "laguerre": [(lambda *x: True, "LaguerreL")],
    "assoc_laguerre": [(lambda *x: True, "LaguerreL")],
    "hermite": [(lambda *x: True, "HermiteH")],
    "jacobi": [(lambda *x: True, "JacobiP")],
    "gegenbauer": [(lambda *x: True, "GegenbauerC")],
    "chebyshevt": [(lambda *x: True, "ChebyshevT")],
    "chebyshevu": [(lambda *x: True, "ChebyshevU")],
    "legendre": [(lambda *x: True, "LegendreP")],
    "assoc_legendre": [(lambda *x: True, "LegendreP")],
    "mathieuc": [(lambda *x: True, "MathieuC")],
    "mathieus": [(lambda *x: True, "MathieuS")],
    "mathieucprime": [(lambda *x: True, "MathieuCPrime")],
    "mathieusprime": [(lambda *x: True, "MathieuSPrime")],
    "stieltjes": [(lambda x: True, "StieltjesGamma")],
    "elliptic_e": [(lambda *x: True, "EllipticE")],
    "elliptic_f": [(lambda *x: True, "EllipticE")],
    "elliptic_k": [(lambda x: True, "EllipticK")],
    "elliptic_pi": [(lambda *x: True, "EllipticPi")],
    "zeta": [(lambda *x: True, "Zeta")],
    "dirichlet_eta": [(lambda x: True, "DirichletEta")],
    "riemann_xi": [(lambda x: True, "RiemannXi")],
    "besseli": [(lambda *x: True, "BesselI")],
    "besselj": [(lambda *x: True, "BesselJ")],
    "besselk": [(lambda *x: True, "BesselK")],
    "bessely": [(lambda *x: True, "BesselY")],
    "hankel1": [(lambda *x: True, "HankelH1")],
    "hankel2": [(lambda *x: True, "HankelH2")],
    "airyai": [(lambda x: True, "AiryAi")],
    "airybi": [(lambda x: True, "AiryBi")],
    "airyaiprime": [(lambda x: True, "AiryAiPrime")],
    "airybiprime": [(lambda x: True, "AiryBiPrime")],
    "polylog": [(lambda *x: True, "PolyLog")],
    "lerchphi": [(lambda *x: True, "LerchPhi")],
    "gcd": [(lambda *x: True, "GCD")],
    "lcm": [(lambda *x: True, "LCM")],
    "jn": [(lambda *x: True, "SphericalBesselJ")],
    "yn": [(lambda *x: True, "SphericalBesselY")],
    "hyper": [(lambda *x: True, "HypergeometricPFQ")],
    "meijerg": [(lambda *x: True, "MeijerG")],
    "appellf1": [(lambda *x: True, "AppellF1")],
    "DiracDelta": [(lambda x: True, "DiracDelta")],
    "Heaviside": [(lambda x: True, "HeavisideTheta")],
    "KroneckerDelta": [(lambda *x: True, "KroneckerDelta")],
    "sqrt": [(lambda x: True, "Sqrt")],  # For automatic rewrites
}


class MCodePrinter(CodePrinter):
    """A printer to convert Python expressions to
    strings of the Wolfram's Mathematica code
    """
    printmethod = "_mcode"
    language = "Wolfram Language"

    _default_settings: dict[str, Any] = dict(CodePrinter._default_settings, **{
        'precision': 15,
        'user_functions': {},
    })

    _number_symbols: set[tuple[Expr, Float]] = set()
    _not_supported: set[Basic] = set()

    def __init__(self, settings={}):
        """Register function mappings supplied by user"""
        CodePrinter.__init__(self, settings)
        self.known_functions = dict(known_functions)
        userfuncs = settings.get('user_functions', {}).copy()
        for k, v in userfuncs.items():
            if not isinstance(v, list):
                userfuncs[k] = [(lambda *x: True, v)]
        self.known_functions.update(userfuncs)

    def _format_code(self, lines):
        return lines

    def _print_Pow(self, expr):
        PREC = precedence(expr)
        return '%s^%s' % (self.parenthesize(expr.base, PREC),
                          self.parenthesize(expr.exp, PREC))

    def _print_Mul(self, expr):
        PREC = precedence(expr)
        c, nc = expr.args_cnc()
        res = super()._print_Mul(expr.func(*c))
        if nc:
            res += '*'
            res += '**'.join(self.parenthesize(a, PREC) for a in nc)
        return res

    def _print_Relational(self, expr):
        lhs_code = self._print(expr.lhs)
        rhs_code = self._print(expr.rhs)
        op = expr.rel_op
        return "{} {} {}".format(lhs_code, op, rhs_code)

    # Primitive numbers
    def _print_Zero(self, expr):
        return '0'

    def _print_One(self, expr):
        return '1'

    def _print_NegativeOne(self, expr):
        return '-1'

    def _print_Half(self, expr):
        return '1/2'

    def _print_ImaginaryUnit(self, expr):
        return 'I'


    # Infinity and invalid numbers
    def _print_Infinity(self, expr):
        return 'Infinity'

    def _print_NegativeInfinity(self, expr):
        return '-Infinity'

    def _print_ComplexInfinity(self, expr):
        return 'ComplexInfinity'

    def _print_NaN(self, expr):
        return 'Indeterminate'


    # Mathematical constants
    def _print_Exp1(self, expr):
        return 'E'

    def _print_Pi(self, expr):
        return 'Pi'

    def _print_GoldenRatio(self, expr):
        return 'GoldenRatio'

    def _print_TribonacciConstant(self, expr):
        expanded = expr.expand(func=True)
        PREC = precedence(expr)
        return self.parenthesize(expanded, PREC)

    def _print_EulerGamma(self, expr):
        return 'EulerGamma'

    def _print_Catalan(self, expr):
        return 'Catalan'


    def _print_list(self, expr):
        return '{' + ', '.join(self.doprint(a) for a in expr) + '}'
    _print_tuple = _print_list
    _print_Tuple = _print_list

    def _print_ImmutableDenseMatrix(self, expr):
        return self.doprint(expr.tolist())

    def _print_ImmutableSparseMatrix(self, expr):

        def print_rule(pos, val):
            return '{} -> {}'.format(
            self.doprint((pos[0]+1, pos[1]+1)), self.doprint(val))

        def print_data():
            items = sorted(expr.todok().items(), key=default_sort_key)
            return '{' + \
                ', '.join(print_rule(k, v) for k, v in items) + \
                '}'

        def print_dims():
            return self.doprint(expr.shape)

        return 'SparseArray[{}, {}]'.format(print_data(), print_dims())

    def _print_ImmutableDenseNDimArray(self, expr):
        return self.doprint(expr.tolist())

    def _print_ImmutableSparseNDimArray(self, expr):
        def print_string_list(string_list):
            return '{' + ', '.join(a for a in string_list) + '}'

        def to_mathematica_index(*args):
            """Helper function to change Python style indexing to
            Pathematica indexing.

            Python indexing (0, 1 ... n-1)
            -> Mathematica indexing (1, 2 ... n)
            """
            return tuple(i + 1 for i in args)

        def print_rule(pos, val):
            """Helper function to print a rule of Mathematica"""
            return '{} -> {}'.format(self.doprint(pos), self.doprint(val))

        def print_data():
            """Helper function to print data part of Mathematica
            sparse array.

            It uses the fourth notation ``SparseArray[data,{d1,d2,...}]``
            from
            https://reference.wolfram.com/language/ref/SparseArray.html

            ``data`` must be formatted with rule.
            """
            return print_string_list(
                [print_rule(
                    to_mathematica_index(*(expr._get_tuple_index(key))),
                    value)
                for key, value in sorted(expr._sparse_array.items())]
            )

        def print_dims():
            """Helper function to print dimensions part of Mathematica
            sparse array.

            It uses the fourth notation ``SparseArray[data,{d1,d2,...}]``
            from
            https://reference.wolfram.com/language/ref/SparseArray.html
            """
            return self.doprint(expr.shape)

        return 'SparseArray[{}, {}]'.format(print_data(), print_dims())

    def _print_Function(self, expr):
        if expr.func.__name__ in self.known_functions:
            cond_mfunc = self.known_functions[expr.func.__name__]
            for cond, mfunc in cond_mfunc:
                if cond(*expr.args):
                    return "%s[%s]" % (mfunc, self.stringify(expr.args, ", "))
        elif expr.func.__name__ in self._rewriteable_functions:
            # Simple rewrite to supported function possible
            target_f, required_fs = self._rewriteable_functions[expr.func.__name__]
            if self._can_print(target_f) and all(self._can_print(f) for f in required_fs):
                return self._print(expr.rewrite(target_f))
        return expr.func.__name__ + "[%s]" % self.stringify(expr.args, ", ")

    _print_MinMaxBase = _print_Function

    def _print_LambertW(self, expr):
        if len(expr.args) == 1:
            return "ProductLog[{}]".format(self._print(expr.args[0]))
        return "ProductLog[{}, {}]".format(
            self._print(expr.args[1]), self._print(expr.args[0]))

    def _print_atan2(self, expr):
        return "ArcTan[{}, {}]".format(
            self._print(expr.args[1]), self._print(expr.args[0]))

    def _print_Integral(self, expr):
        if len(expr.variables) == 1 and not expr.limits[0][1:]:
            args = [expr.args[0], expr.variables[0]]
        else:
            args = expr.args
        return "Hold[Integrate[" + ', '.join(self.doprint(a) for a in args) + "]]"

    def _print_Sum(self, expr):
        return "Hold[Sum[" + ', '.join(self.doprint(a) for a in expr.args) + "]]"

    def _print_Derivative(self, expr):
        dexpr = expr.expr
        dvars = [i[0] if i[1] == 1 else i for i in expr.variable_count]
        return "Hold[D[" + ', '.join(self.doprint(a) for a in [dexpr] + dvars) + "]]"


    def _get_comment(self, text):
        return "(* {} *)".format(text)


def mathematica_code(expr, **settings):
    r"""Converts an expr to a string of the Wolfram Mathematica code

    Examples
    ========

    >>> from sympy import mathematica_code as mcode, symbols, sin
    >>> x = symbols('x')
    >>> mcode(sin(x).series(x).removeO())
    '(1/120)*x^5 - 1/6*x^3 + x'
    """
    return MCodePrinter(settings).doprint(expr)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\assumptions\ask_generated.py ===
"""
Do NOT manually edit this file.
Instead, run ./bin/ask_update.py.
"""

from sympy.assumptions.ask import Q
from sympy.assumptions.cnf import Literal
from sympy.core.cache import cacheit

@cacheit
def get_all_known_facts():
    """
    Known facts between unary predicates as CNF clauses.
    """
    return {
        frozenset((Literal(Q.algebraic, False), Literal(Q.imaginary, True), Literal(Q.transcendental, False))),
        frozenset((Literal(Q.algebraic, False), Literal(Q.negative, True), Literal(Q.transcendental, False))),
        frozenset((Literal(Q.algebraic, False), Literal(Q.positive, True), Literal(Q.transcendental, False))),
        frozenset((Literal(Q.algebraic, False), Literal(Q.rational, True))),
        frozenset((Literal(Q.algebraic, False), Literal(Q.transcendental, False), Literal(Q.zero, True))),
        frozenset((Literal(Q.algebraic, True), Literal(Q.finite, False))),
        frozenset((Literal(Q.algebraic, True), Literal(Q.transcendental, True))),
        frozenset((Literal(Q.antihermitian, False), Literal(Q.hermitian, False), Literal(Q.zero, True))),
        frozenset((Literal(Q.antihermitian, False), Literal(Q.imaginary, True))),
        frozenset((Literal(Q.commutative, False), Literal(Q.finite, True))),
        frozenset((Literal(Q.commutative, False), Literal(Q.infinite, True))),
        frozenset((Literal(Q.complex_elements, False), Literal(Q.real_elements, True))),
        frozenset((Literal(Q.composite, False), Literal(Q.even, True), Literal(Q.positive, True), Literal(Q.prime, False))),
        frozenset((Literal(Q.composite, True), Literal(Q.even, False), Literal(Q.odd, False))),
        frozenset((Literal(Q.composite, True), Literal(Q.positive, False))),
        frozenset((Literal(Q.composite, True), Literal(Q.prime, True))),
        frozenset((Literal(Q.diagonal, False), Literal(Q.lower_triangular, True), Literal(Q.upper_triangular, True))),
        frozenset((Literal(Q.diagonal, True), Literal(Q.lower_triangular, False))),
        frozenset((Literal(Q.diagonal, True), Literal(Q.normal, False))),
        frozenset((Literal(Q.diagonal, True), Literal(Q.symmetric, False))),
        frozenset((Literal(Q.diagonal, True), Literal(Q.upper_triangular, False))),
        frozenset((Literal(Q.even, False), Literal(Q.odd, False), Literal(Q.prime, True))),
        frozenset((Literal(Q.even, False), Literal(Q.zero, True))),
        frozenset((Literal(Q.even, True), Literal(Q.odd, True))),
        frozenset((Literal(Q.even, True), Literal(Q.rational, False))),
        frozenset((Literal(Q.finite, False), Literal(Q.transcendental, True))),
        frozenset((Literal(Q.finite, True), Literal(Q.infinite, True))),
        frozenset((Literal(Q.fullrank, False), Literal(Q.invertible, True))),
        frozenset((Literal(Q.fullrank, True), Literal(Q.invertible, False), Literal(Q.square, True))),
        frozenset((Literal(Q.hermitian, False), Literal(Q.negative, True))),
        frozenset((Literal(Q.hermitian, False), Literal(Q.positive, True))),
        frozenset((Literal(Q.hermitian, False), Literal(Q.zero, True))),
        frozenset((Literal(Q.imaginary, True), Literal(Q.negative, True))),
        frozenset((Literal(Q.imaginary, True), Literal(Q.positive, True))),
        frozenset((Literal(Q.imaginary, True), Literal(Q.zero, True))),
        frozenset((Literal(Q.infinite, False), Literal(Q.negative_infinite, True))),
        frozenset((Literal(Q.infinite, False), Literal(Q.positive_infinite, True))),
        frozenset((Literal(Q.integer_elements, True), Literal(Q.real_elements, False))),
        frozenset((Literal(Q.invertible, False), Literal(Q.positive_definite, True))),
        frozenset((Literal(Q.invertible, False), Literal(Q.singular, False))),
        frozenset((Literal(Q.invertible, False), Literal(Q.unitary, True))),
        frozenset((Literal(Q.invertible, True), Literal(Q.singular, True))),
        frozenset((Literal(Q.invertible, True), Literal(Q.square, False))),
        frozenset((Literal(Q.irrational, False), Literal(Q.negative, True), Literal(Q.rational, False))),
        frozenset((Literal(Q.irrational, False), Literal(Q.positive, True), Literal(Q.rational, False))),
        frozenset((Literal(Q.irrational, False), Literal(Q.rational, False), Literal(Q.zero, True))),
        frozenset((Literal(Q.irrational, True), Literal(Q.negative, False), Literal(Q.positive, False), Literal(Q.zero, False))),
        frozenset((Literal(Q.irrational, True), Literal(Q.rational, True))),
        frozenset((Literal(Q.lower_triangular, False), Literal(Q.triangular, True), Literal(Q.upper_triangular, False))),
        frozenset((Literal(Q.lower_triangular, True), Literal(Q.triangular, False))),
        frozenset((Literal(Q.negative, False), Literal(Q.positive, False), Literal(Q.rational, True), Literal(Q.zero, False))),
        frozenset((Literal(Q.negative, True), Literal(Q.negative_infinite, True))),
        frozenset((Literal(Q.negative, True), Literal(Q.positive, True))),
        frozenset((Literal(Q.negative, True), Literal(Q.positive_infinite, True))),
        frozenset((Literal(Q.negative, True), Literal(Q.zero, True))),
        frozenset((Literal(Q.negative_infinite, True), Literal(Q.positive, True))),
        frozenset((Literal(Q.negative_infinite, True), Literal(Q.positive_infinite, True))),
        frozenset((Literal(Q.negative_infinite, True), Literal(Q.zero, True))),
        frozenset((Literal(Q.normal, False), Literal(Q.unitary, True))),
        frozenset((Literal(Q.normal, True), Literal(Q.square, False))),
        frozenset((Literal(Q.odd, True), Literal(Q.rational, False))),
        frozenset((Literal(Q.orthogonal, False), Literal(Q.real_elements, True), Literal(Q.unitary, True))),
        frozenset((Literal(Q.orthogonal, True), Literal(Q.positive_definite, False))),
        frozenset((Literal(Q.orthogonal, True), Literal(Q.unitary, False))),
        frozenset((Literal(Q.positive, False), Literal(Q.prime, True))),
        frozenset((Literal(Q.positive, True), Literal(Q.positive_infinite, True))),
        frozenset((Literal(Q.positive, True), Literal(Q.zero, True))),
        frozenset((Literal(Q.positive_infinite, True), Literal(Q.zero, True))),
        frozenset((Literal(Q.square, False), Literal(Q.symmetric, True))),
        frozenset((Literal(Q.triangular, False), Literal(Q.unit_triangular, True))),
        frozenset((Literal(Q.triangular, False), Literal(Q.upper_triangular, True)))
    }

@cacheit
def get_all_known_matrix_facts():
    """
    Known facts between unary predicates for matrices as CNF clauses.
    """
    return {
        frozenset((Literal(Q.complex_elements, False), Literal(Q.real_elements, True))),
        frozenset((Literal(Q.diagonal, False), Literal(Q.lower_triangular, True), Literal(Q.upper_triangular, True))),
        frozenset((Literal(Q.diagonal, True), Literal(Q.lower_triangular, False))),
        frozenset((Literal(Q.diagonal, True), Literal(Q.normal, False))),
        frozenset((Literal(Q.diagonal, True), Literal(Q.symmetric, False))),
        frozenset((Literal(Q.diagonal, True), Literal(Q.upper_triangular, False))),
        frozenset((Literal(Q.fullrank, False), Literal(Q.invertible, True))),
        frozenset((Literal(Q.fullrank, True), Literal(Q.invertible, False), Literal(Q.square, True))),
        frozenset((Literal(Q.integer_elements, True), Literal(Q.real_elements, False))),
        frozenset((Literal(Q.invertible, False), Literal(Q.positive_definite, True))),
        frozenset((Literal(Q.invertible, False), Literal(Q.singular, False))),
        frozenset((Literal(Q.invertible, False), Literal(Q.unitary, True))),
        frozenset((Literal(Q.invertible, True), Literal(Q.singular, True))),
        frozenset((Literal(Q.invertible, True), Literal(Q.square, False))),
        frozenset((Literal(Q.lower_triangular, False), Literal(Q.triangular, True), Literal(Q.upper_triangular, False))),
        frozenset((Literal(Q.lower_triangular, True), Literal(Q.triangular, False))),
        frozenset((Literal(Q.normal, False), Literal(Q.unitary, True))),
        frozenset((Literal(Q.normal, True), Literal(Q.square, False))),
        frozenset((Literal(Q.orthogonal, False), Literal(Q.real_elements, True), Literal(Q.unitary, True))),
        frozenset((Literal(Q.orthogonal, True), Literal(Q.positive_definite, False))),
        frozenset((Literal(Q.orthogonal, True), Literal(Q.unitary, False))),
        frozenset((Literal(Q.square, False), Literal(Q.symmetric, True))),
        frozenset((Literal(Q.triangular, False), Literal(Q.unit_triangular, True))),
        frozenset((Literal(Q.triangular, False), Literal(Q.upper_triangular, True)))
    }

@cacheit
def get_all_known_number_facts():
    """
    Known facts between unary predicates for numbers as CNF clauses.
    """
    return {
        frozenset((Literal(Q.algebraic, False), Literal(Q.imaginary, True), Literal(Q.transcendental, False))),
        frozenset((Literal(Q.algebraic, False), Literal(Q.negative, True), Literal(Q.transcendental, False))),
        frozenset((Literal(Q.algebraic, False), Literal(Q.positive, True), Literal(Q.transcendental, False))),
        frozenset((Literal(Q.algebraic, False), Literal(Q.rational, True))),
        frozenset((Literal(Q.algebraic, False), Literal(Q.transcendental, False), Literal(Q.zero, True))),
        frozenset((Literal(Q.algebraic, True), Literal(Q.finite, False))),
        frozenset((Literal(Q.algebraic, True), Literal(Q.transcendental, True))),
        frozenset((Literal(Q.antihermitian, False), Literal(Q.hermitian, False), Literal(Q.zero, True))),
        frozenset((Literal(Q.antihermitian, False), Literal(Q.imaginary, True))),
        frozenset((Literal(Q.commutative, False), Literal(Q.finite, True))),
        frozenset((Literal(Q.commutative, False), Literal(Q.infinite, True))),
        frozenset((Literal(Q.composite, False), Literal(Q.even, True), Literal(Q.positive, True), Literal(Q.prime, False))),
        frozenset((Literal(Q.composite, True), Literal(Q.even, False), Literal(Q.odd, False))),
        frozenset((Literal(Q.composite, True), Literal(Q.positive, False))),
        frozenset((Literal(Q.composite, True), Literal(Q.prime, True))),
        frozenset((Literal(Q.even, False), Literal(Q.odd, False), Literal(Q.prime, True))),
        frozenset((Literal(Q.even, False), Literal(Q.zero, True))),
        frozenset((Literal(Q.even, True), Literal(Q.odd, True))),
        frozenset((Literal(Q.even, True), Literal(Q.rational, False))),
        frozenset((Literal(Q.finite, False), Literal(Q.transcendental, True))),
        frozenset((Literal(Q.finite, True), Literal(Q.infinite, True))),
        frozenset((Literal(Q.hermitian, False), Literal(Q.negative, True))),
        frozenset((Literal(Q.hermitian, False), Literal(Q.positive, True))),
        frozenset((Literal(Q.hermitian, False), Literal(Q.zero, True))),
        frozenset((Literal(Q.imaginary, True), Literal(Q.negative, True))),
        frozenset((Literal(Q.imaginary, True), Literal(Q.positive, True))),
        frozenset((Literal(Q.imaginary, True), Literal(Q.zero, True))),
        frozenset((Literal(Q.infinite, False), Literal(Q.negative_infinite, True))),
        frozenset((Literal(Q.infinite, False), Literal(Q.positive_infinite, True))),
        frozenset((Literal(Q.irrational, False), Literal(Q.negative, True), Literal(Q.rational, False))),
        frozenset((Literal(Q.irrational, False), Literal(Q.positive, True), Literal(Q.rational, False))),
        frozenset((Literal(Q.irrational, False), Literal(Q.rational, False), Literal(Q.zero, True))),
        frozenset((Literal(Q.irrational, True), Literal(Q.negative, False), Literal(Q.positive, False), Literal(Q.zero, False))),
        frozenset((Literal(Q.irrational, True), Literal(Q.rational, True))),
        frozenset((Literal(Q.negative, False), Literal(Q.positive, False), Literal(Q.rational, True), Literal(Q.zero, False))),
        frozenset((Literal(Q.negative, True), Literal(Q.negative_infinite, True))),
        frozenset((Literal(Q.negative, True), Literal(Q.positive, True))),
        frozenset((Literal(Q.negative, True), Literal(Q.positive_infinite, True))),
        frozenset((Literal(Q.negative, True), Literal(Q.zero, True))),
        frozenset((Literal(Q.negative_infinite, True), Literal(Q.positive, True))),
        frozenset((Literal(Q.negative_infinite, True), Literal(Q.positive_infinite, True))),
        frozenset((Literal(Q.negative_infinite, True), Literal(Q.zero, True))),
        frozenset((Literal(Q.odd, True), Literal(Q.rational, False))),
        frozenset((Literal(Q.positive, False), Literal(Q.prime, True))),
        frozenset((Literal(Q.positive, True), Literal(Q.positive_infinite, True))),
        frozenset((Literal(Q.positive, True), Literal(Q.zero, True))),
        frozenset((Literal(Q.positive_infinite, True), Literal(Q.zero, True)))
    }

@cacheit
def get_known_facts_dict():
    """
    Logical relations between unary predicates as dictionary.

    Each key is a predicate, and item is two groups of predicates.
    First group contains the predicates which are implied by the key, and
    second group contains the predicates which are rejected by the key.

    """
    return {
        Q.algebraic: (set([Q.algebraic, Q.commutative, Q.complex, Q.finite]),
        set([Q.infinite, Q.negative_infinite, Q.positive_infinite,
        Q.transcendental])),
        Q.antihermitian: (set([Q.antihermitian]), set([])),
        Q.commutative: (set([Q.commutative]), set([])),
        Q.complex: (set([Q.commutative, Q.complex, Q.finite]),
        set([Q.infinite, Q.negative_infinite, Q.positive_infinite])),
        Q.complex_elements: (set([Q.complex_elements]), set([])),
        Q.composite: (set([Q.algebraic, Q.commutative, Q.complex, Q.composite,
        Q.extended_nonnegative, Q.extended_nonzero,
        Q.extended_positive, Q.extended_real, Q.finite, Q.hermitian,
        Q.integer, Q.nonnegative, Q.nonzero, Q.positive, Q.rational,
        Q.real]), set([Q.extended_negative, Q.extended_nonpositive,
        Q.imaginary, Q.infinite, Q.irrational, Q.negative,
        Q.negative_infinite, Q.nonpositive, Q.positive_infinite,
        Q.prime, Q.transcendental, Q.zero])),
        Q.diagonal: (set([Q.diagonal, Q.lower_triangular, Q.normal, Q.square,
        Q.symmetric, Q.triangular, Q.upper_triangular]), set([])),
        Q.even: (set([Q.algebraic, Q.commutative, Q.complex, Q.even,
        Q.extended_real, Q.finite, Q.hermitian, Q.integer, Q.rational,
        Q.real]), set([Q.imaginary, Q.infinite, Q.irrational,
        Q.negative_infinite, Q.odd, Q.positive_infinite,
        Q.transcendental])),
        Q.extended_negative: (set([Q.commutative, Q.extended_negative,
        Q.extended_nonpositive, Q.extended_nonzero, Q.extended_real]),
        set([Q.composite, Q.extended_nonnegative, Q.extended_positive,
        Q.imaginary, Q.nonnegative, Q.positive, Q.positive_infinite,
        Q.prime, Q.zero])),
        Q.extended_nonnegative: (set([Q.commutative, Q.extended_nonnegative,
        Q.extended_real]), set([Q.extended_negative, Q.imaginary,
        Q.negative, Q.negative_infinite])),
        Q.extended_nonpositive: (set([Q.commutative, Q.extended_nonpositive,
        Q.extended_real]), set([Q.composite, Q.extended_positive,
        Q.imaginary, Q.positive, Q.positive_infinite, Q.prime])),
        Q.extended_nonzero: (set([Q.commutative, Q.extended_nonzero,
        Q.extended_real]), set([Q.imaginary, Q.zero])),
        Q.extended_positive: (set([Q.commutative, Q.extended_nonnegative,
        Q.extended_nonzero, Q.extended_positive, Q.extended_real]),
        set([Q.extended_negative, Q.extended_nonpositive, Q.imaginary,
        Q.negative, Q.negative_infinite, Q.nonpositive, Q.zero])),
        Q.extended_real: (set([Q.commutative, Q.extended_real]),
        set([Q.imaginary])),
        Q.finite: (set([Q.commutative, Q.finite]), set([Q.infinite,
        Q.negative_infinite, Q.positive_infinite])),
        Q.fullrank: (set([Q.fullrank]), set([])),
        Q.hermitian: (set([Q.hermitian]), set([])),
        Q.imaginary: (set([Q.antihermitian, Q.commutative, Q.complex,
        Q.finite, Q.imaginary]), set([Q.composite, Q.even,
        Q.extended_negative, Q.extended_nonnegative,
        Q.extended_nonpositive, Q.extended_nonzero,
        Q.extended_positive, Q.extended_real, Q.infinite, Q.integer,
        Q.irrational, Q.negative, Q.negative_infinite, Q.nonnegative,
        Q.nonpositive, Q.nonzero, Q.odd, Q.positive,
        Q.positive_infinite, Q.prime, Q.rational, Q.real, Q.zero])),
        Q.infinite: (set([Q.commutative, Q.infinite]), set([Q.algebraic,
        Q.complex, Q.composite, Q.even, Q.finite, Q.imaginary,
        Q.integer, Q.irrational, Q.negative, Q.nonnegative,
        Q.nonpositive, Q.nonzero, Q.odd, Q.positive, Q.prime,
        Q.rational, Q.real, Q.transcendental, Q.zero])),
        Q.integer: (set([Q.algebraic, Q.commutative, Q.complex,
        Q.extended_real, Q.finite, Q.hermitian, Q.integer, Q.rational,
        Q.real]), set([Q.imaginary, Q.infinite, Q.irrational,
        Q.negative_infinite, Q.positive_infinite, Q.transcendental])),
        Q.integer_elements: (set([Q.complex_elements, Q.integer_elements,
        Q.real_elements]), set([])),
        Q.invertible: (set([Q.fullrank, Q.invertible, Q.square]),
        set([Q.singular])),
        Q.irrational: (set([Q.commutative, Q.complex, Q.extended_nonzero,
        Q.extended_real, Q.finite, Q.hermitian, Q.irrational,
        Q.nonzero, Q.real]), set([Q.composite, Q.even, Q.imaginary,
        Q.infinite, Q.integer, Q.negative_infinite, Q.odd,
        Q.positive_infinite, Q.prime, Q.rational, Q.zero])),
        Q.is_true: (set([Q.is_true]), set([])),
        Q.lower_triangular: (set([Q.lower_triangular, Q.triangular]), set([])),
        Q.negative: (set([Q.commutative, Q.complex, Q.extended_negative,
        Q.extended_nonpositive, Q.extended_nonzero, Q.extended_real,
        Q.finite, Q.hermitian, Q.negative, Q.nonpositive, Q.nonzero,
        Q.real]), set([Q.composite, Q.extended_nonnegative,
        Q.extended_positive, Q.imaginary, Q.infinite,
        Q.negative_infinite, Q.nonnegative, Q.positive,
        Q.positive_infinite, Q.prime, Q.zero])),
        Q.negative_infinite: (set([Q.commutative, Q.extended_negative,
        Q.extended_nonpositive, Q.extended_nonzero, Q.extended_real,
        Q.infinite, Q.negative_infinite]), set([Q.algebraic,
        Q.complex, Q.composite, Q.even, Q.extended_nonnegative,
        Q.extended_positive, Q.finite, Q.imaginary, Q.integer,
        Q.irrational, Q.negative, Q.nonnegative, Q.nonpositive,
        Q.nonzero, Q.odd, Q.positive, Q.positive_infinite, Q.prime,
        Q.rational, Q.real, Q.transcendental, Q.zero])),
        Q.noninteger: (set([Q.noninteger]), set([])),
        Q.nonnegative: (set([Q.commutative, Q.complex, Q.extended_nonnegative,
        Q.extended_real, Q.finite, Q.hermitian, Q.nonnegative,
        Q.real]), set([Q.extended_negative, Q.imaginary, Q.infinite,
        Q.negative, Q.negative_infinite, Q.positive_infinite])),
        Q.nonpositive: (set([Q.commutative, Q.complex, Q.extended_nonpositive,
        Q.extended_real, Q.finite, Q.hermitian, Q.nonpositive,
        Q.real]), set([Q.composite, Q.extended_positive, Q.imaginary,
        Q.infinite, Q.negative_infinite, Q.positive,
        Q.positive_infinite, Q.prime])),
        Q.nonzero: (set([Q.commutative, Q.complex, Q.extended_nonzero,
        Q.extended_real, Q.finite, Q.hermitian, Q.nonzero, Q.real]),
        set([Q.imaginary, Q.infinite, Q.negative_infinite,
        Q.positive_infinite, Q.zero])),
        Q.normal: (set([Q.normal, Q.square]), set([])),
        Q.odd: (set([Q.algebraic, Q.commutative, Q.complex,
        Q.extended_nonzero, Q.extended_real, Q.finite, Q.hermitian,
        Q.integer, Q.nonzero, Q.odd, Q.rational, Q.real]),
        set([Q.even, Q.imaginary, Q.infinite, Q.irrational,
        Q.negative_infinite, Q.positive_infinite, Q.transcendental,
        Q.zero])),
        Q.orthogonal: (set([Q.fullrank, Q.invertible, Q.normal, Q.orthogonal,
        Q.positive_definite, Q.square, Q.unitary]), set([Q.singular])),
        Q.positive: (set([Q.commutative, Q.complex, Q.extended_nonnegative,
        Q.extended_nonzero, Q.extended_positive, Q.extended_real,
        Q.finite, Q.hermitian, Q.nonnegative, Q.nonzero, Q.positive,
        Q.real]), set([Q.extended_negative, Q.extended_nonpositive,
        Q.imaginary, Q.infinite, Q.negative, Q.negative_infinite,
        Q.nonpositive, Q.positive_infinite, Q.zero])),
        Q.positive_definite: (set([Q.fullrank, Q.invertible,
        Q.positive_definite, Q.square]), set([Q.singular])),
        Q.positive_infinite: (set([Q.commutative, Q.extended_nonnegative,
        Q.extended_nonzero, Q.extended_positive, Q.extended_real,
        Q.infinite, Q.positive_infinite]), set([Q.algebraic,
        Q.complex, Q.composite, Q.even, Q.extended_negative,
        Q.extended_nonpositive, Q.finite, Q.imaginary, Q.integer,
        Q.irrational, Q.negative, Q.negative_infinite, Q.nonnegative,
        Q.nonpositive, Q.nonzero, Q.odd, Q.positive, Q.prime,
        Q.rational, Q.real, Q.transcendental, Q.zero])),
        Q.prime: (set([Q.algebraic, Q.commutative, Q.complex,
        Q.extended_nonnegative, Q.extended_nonzero,
        Q.extended_positive, Q.extended_real, Q.finite, Q.hermitian,
        Q.integer, Q.nonnegative, Q.nonzero, Q.positive, Q.prime,
        Q.rational, Q.real]), set([Q.composite, Q.extended_negative,
        Q.extended_nonpositive, Q.imaginary, Q.infinite, Q.irrational,
        Q.negative, Q.negative_infinite, Q.nonpositive,
        Q.positive_infinite, Q.transcendental, Q.zero])),
        Q.rational: (set([Q.algebraic, Q.commutative, Q.complex,
        Q.extended_real, Q.finite, Q.hermitian, Q.rational, Q.real]),
        set([Q.imaginary, Q.infinite, Q.irrational,
        Q.negative_infinite, Q.positive_infinite, Q.transcendental])),
        Q.real: (set([Q.commutative, Q.complex, Q.extended_real, Q.finite,
        Q.hermitian, Q.real]), set([Q.imaginary, Q.infinite,
        Q.negative_infinite, Q.positive_infinite])),
        Q.real_elements: (set([Q.complex_elements, Q.real_elements]), set([])),
        Q.singular: (set([Q.singular]), set([Q.invertible, Q.orthogonal,
        Q.positive_definite, Q.unitary])),
        Q.square: (set([Q.square]), set([])),
        Q.symmetric: (set([Q.square, Q.symmetric]), set([])),
        Q.transcendental: (set([Q.commutative, Q.complex, Q.finite,
        Q.transcendental]), set([Q.algebraic, Q.composite, Q.even,
        Q.infinite, Q.integer, Q.negative_infinite, Q.odd,
        Q.positive_infinite, Q.prime, Q.rational, Q.zero])),
        Q.triangular: (set([Q.triangular]), set([])),
        Q.unit_triangular: (set([Q.triangular, Q.unit_triangular]), set([])),
        Q.unitary: (set([Q.fullrank, Q.invertible, Q.normal, Q.square,
        Q.unitary]), set([Q.singular])),
        Q.upper_triangular: (set([Q.triangular, Q.upper_triangular]), set([])),
        Q.zero: (set([Q.algebraic, Q.commutative, Q.complex, Q.even,
        Q.extended_nonnegative, Q.extended_nonpositive,
        Q.extended_real, Q.finite, Q.hermitian, Q.integer,
        Q.nonnegative, Q.nonpositive, Q.rational, Q.real, Q.zero]),
        set([Q.composite, Q.extended_negative, Q.extended_nonzero,
        Q.extended_positive, Q.imaginary, Q.infinite, Q.irrational,
        Q.negative, Q.negative_infinite, Q.nonzero, Q.odd, Q.positive,
        Q.positive_infinite, Q.prime, Q.transcendental])),
    }

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\simplify\epathtools.py ===
"""Tools for manipulation of expressions using paths. """

from sympy.core import Basic


class EPath:
    r"""
    Manipulate expressions using paths.

    EPath grammar in EBNF notation::

        literal   ::= /[A-Za-z_][A-Za-z_0-9]*/
        number    ::= /-?\d+/
        type      ::= literal
        attribute ::= literal "?"
        all       ::= "*"
        slice     ::= "[" number? (":" number? (":" number?)?)? "]"
        range     ::= all | slice
        query     ::= (type | attribute) ("|" (type | attribute))*
        selector  ::= range | query range?
        path      ::= "/" selector ("/" selector)*

    See the docstring of the epath() function.

    """

    __slots__ = ("_path", "_epath")

    def __new__(cls, path):
        """Construct new EPath. """
        if isinstance(path, EPath):
            return path

        if not path:
            raise ValueError("empty EPath")

        _path = path

        if path[0] == '/':
            path = path[1:]
        else:
            raise NotImplementedError("non-root EPath")

        epath = []

        for selector in path.split('/'):
            selector = selector.strip()

            if not selector:
                raise ValueError("empty selector")

            index = 0

            for c in selector:
                if c.isalnum() or c in ('_', '|', '?'):
                    index += 1
                else:
                    break

            attrs = []
            types = []

            if index:
                elements = selector[:index]
                selector = selector[index:]

                for element in elements.split('|'):
                    element = element.strip()

                    if not element:
                        raise ValueError("empty element")

                    if element.endswith('?'):
                        attrs.append(element[:-1])
                    else:
                        types.append(element)

            span = None

            if selector == '*':
                pass
            else:
                if selector.startswith('['):
                    try:
                        i = selector.index(']')
                    except ValueError:
                        raise ValueError("expected ']', got EOL")

                    _span, span = selector[1:i], []

                    if ':' not in _span:
                        span = int(_span)
                    else:
                        for elt in _span.split(':', 3):
                            if not elt:
                                span.append(None)
                            else:
                                span.append(int(elt))

                        span = slice(*span)

                    selector = selector[i + 1:]

                if selector:
                    raise ValueError("trailing characters in selector")

            epath.append((attrs, types, span))

        obj = object.__new__(cls)

        obj._path = _path
        obj._epath = epath

        return obj

    def __repr__(self):
        return "%s(%r)" % (self.__class__.__name__, self._path)

    def _get_ordered_args(self, expr):
        """Sort ``expr.args`` using printing order. """
        if expr.is_Add:
            return expr.as_ordered_terms()
        elif expr.is_Mul:
            return expr.as_ordered_factors()
        else:
            return expr.args

    def _hasattrs(self, expr, attrs) -> bool:
        """Check if ``expr`` has any of ``attrs``. """
        return all(hasattr(expr, attr) for attr in attrs)

    def _hastypes(self, expr, types):
        """Check if ``expr`` is any of ``types``. """
        _types = [ cls.__name__ for cls in expr.__class__.mro() ]
        return bool(set(_types).intersection(types))

    def _has(self, expr, attrs, types):
        """Apply ``_hasattrs`` and ``_hastypes`` to ``expr``. """
        if not (attrs or types):
            return True

        if attrs and self._hasattrs(expr, attrs):
            return True

        if types and self._hastypes(expr, types):
            return True

        return False

    def apply(self, expr, func, args=None, kwargs=None):
        """
        Modify parts of an expression selected by a path.

        Examples
        ========

        >>> from sympy.simplify.epathtools import EPath
        >>> from sympy import sin, cos, E
        >>> from sympy.abc import x, y, z, t

        >>> path = EPath("/*/[0]/Symbol")
        >>> expr = [((x, 1), 2), ((3, y), z)]

        >>> path.apply(expr, lambda expr: expr**2)
        [((x**2, 1), 2), ((3, y**2), z)]

        >>> path = EPath("/*/*/Symbol")
        >>> expr = t + sin(x + 1) + cos(x + y + E)

        >>> path.apply(expr, lambda expr: 2*expr)
        t + sin(2*x + 1) + cos(2*x + 2*y + E)

        """
        def _apply(path, expr, func):
            if not path:
                return func(expr)
            else:
                selector, path = path[0], path[1:]
                attrs, types, span = selector

                if isinstance(expr, Basic):
                    if not expr.is_Atom:
                        args, basic = self._get_ordered_args(expr), True
                    else:
                        return expr
                elif hasattr(expr, '__iter__'):
                    args, basic = expr, False
                else:
                    return expr

                args = list(args)

                if span is not None:
                    if isinstance(span, slice):
                        indices = range(*span.indices(len(args)))
                    else:
                        indices = [span]
                else:
                    indices = range(len(args))

                for i in indices:
                    try:
                        arg = args[i]
                    except IndexError:
                        continue

                    if self._has(arg, attrs, types):
                        args[i] = _apply(path, arg, func)

                if basic:
                    return expr.func(*args)
                else:
                    return expr.__class__(args)

        _args, _kwargs = args or (), kwargs or {}
        _func = lambda expr: func(expr, *_args, **_kwargs)

        return _apply(self._epath, expr, _func)

    def select(self, expr):
        """
        Retrieve parts of an expression selected by a path.

        Examples
        ========

        >>> from sympy.simplify.epathtools import EPath
        >>> from sympy import sin, cos, E
        >>> from sympy.abc import x, y, z, t

        >>> path = EPath("/*/[0]/Symbol")
        >>> expr = [((x, 1), 2), ((3, y), z)]

        >>> path.select(expr)
        [x, y]

        >>> path = EPath("/*/*/Symbol")
        >>> expr = t + sin(x + 1) + cos(x + y + E)

        >>> path.select(expr)
        [x, x, y]

        """
        result = []

        def _select(path, expr):
            if not path:
                result.append(expr)
            else:
                selector, path = path[0], path[1:]
                attrs, types, span = selector

                if isinstance(expr, Basic):
                    args = self._get_ordered_args(expr)
                elif hasattr(expr, '__iter__'):
                    args = expr
                else:
                    return

                if span is not None:
                    if isinstance(span, slice):
                        args = args[span]
                    else:
                        try:
                            args = [args[span]]
                        except IndexError:
                            return

                for arg in args:
                    if self._has(arg, attrs, types):
                        _select(path, arg)

        _select(self._epath, expr)
        return result


def epath(path, expr=None, func=None, args=None, kwargs=None):
    r"""
    Manipulate parts of an expression selected by a path.

    Explanation
    ===========

    This function allows to manipulate large nested expressions in single
    line of code, utilizing techniques to those applied in XML processing
    standards (e.g. XPath).

    If ``func`` is ``None``, :func:`epath` retrieves elements selected by
    the ``path``. Otherwise it applies ``func`` to each matching element.

    Note that it is more efficient to create an EPath object and use the select
    and apply methods of that object, since this will compile the path string
    only once.  This function should only be used as a convenient shortcut for
    interactive use.

    This is the supported syntax:

    * select all: ``/*``
          Equivalent of ``for arg in args:``.
    * select slice: ``/[0]`` or ``/[1:5]`` or ``/[1:5:2]``
          Supports standard Python's slice syntax.
    * select by type: ``/list`` or ``/list|tuple``
          Emulates ``isinstance()``.
    * select by attribute: ``/__iter__?``
          Emulates ``hasattr()``.

    Parameters
    ==========

    path : str | EPath
        A path as a string or a compiled EPath.
    expr : Basic | iterable
        An expression or a container of expressions.
    func : callable (optional)
        A callable that will be applied to matching parts.
    args : tuple (optional)
        Additional positional arguments to ``func``.
    kwargs : dict (optional)
        Additional keyword arguments to ``func``.

    Examples
    ========

    >>> from sympy.simplify.epathtools import epath
    >>> from sympy import sin, cos, E
    >>> from sympy.abc import x, y, z, t

    >>> path = "/*/[0]/Symbol"
    >>> expr = [((x, 1), 2), ((3, y), z)]

    >>> epath(path, expr)
    [x, y]
    >>> epath(path, expr, lambda expr: expr**2)
    [((x**2, 1), 2), ((3, y**2), z)]

    >>> path = "/*/*/Symbol"
    >>> expr = t + sin(x + 1) + cos(x + y + E)

    >>> epath(path, expr)
    [x, x, y]
    >>> epath(path, expr, lambda expr: 2*expr)
    t + sin(2*x + 1) + cos(2*x + 2*y + E)

    """
    _epath = EPath(path)

    if expr is None:
        return _epath
    if func is None:
        return _epath.select(expr)
    else:
        return _epath.apply(expr, func, args, kwargs)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\humanfriendly\usage.py ===
# Human friendly input/output in Python.
#
# Author: Peter Odding <peter@peterodding.com>
# Last Change: June 11, 2021
# URL: https://humanfriendly.readthedocs.io

"""
Parsing and reformatting of usage messages.

The :mod:`~humanfriendly.usage` module parses and reformats usage messages:

- The :func:`format_usage()` function takes a usage message and inserts ANSI
  escape sequences that highlight items of special significance like command
  line options, meta variables, etc. The resulting usage message is (intended
  to be) easier to read on a terminal.

- The :func:`render_usage()` function takes a usage message and rewrites it to
  reStructuredText_ suitable for inclusion in the documentation of a Python
  package. This provides a DRY solution to keeping a single authoritative
  definition of the usage message while making it easily available in
  documentation. As a cherry on the cake it's not just a pre-formatted dump of
  the usage message but a nicely formatted reStructuredText_ fragment.

- The remaining functions in this module support the two functions above.

Usage messages in general are free format of course, however the functions in
this module assume a certain structure from usage messages in order to
successfully parse and reformat them, refer to :func:`parse_usage()` for
details.

.. _DRY: https://en.wikipedia.org/wiki/Don%27t_repeat_yourself
.. _reStructuredText: https://en.wikipedia.org/wiki/ReStructuredText
"""

# Standard library modules.
import csv
import functools
import logging
import re

# Standard library module or external dependency (see setup.py).
from importlib import import_module

# Modules included in our package.
from humanfriendly.compat import StringIO
from humanfriendly.text import dedent, split_paragraphs, trim_empty_lines

# Public identifiers that require documentation.
__all__ = (
    'find_meta_variables',
    'format_usage',
    'import_module',  # previously exported (backwards compatibility)
    'inject_usage',
    'parse_usage',
    'render_usage',
    'USAGE_MARKER',
)

USAGE_MARKER = "Usage:"
"""The string that starts the first line of a usage message."""

START_OF_OPTIONS_MARKER = "Supported options:"
"""The string that marks the start of the documented command line options."""

# Compiled regular expression used to tokenize usage messages.
USAGE_PATTERN = re.compile(r'''
    # Make sure whatever we're matching isn't preceded by a non-whitespace
    # character.
    (?<!\S)
    (
        # A short command line option or a long command line option
        # (possibly including a meta variable for a value).
        (-\w|--\w+(-\w+)*(=\S+)?)
        # Or ...
        |
        # An environment variable.
        \$[A-Za-z_][A-Za-z0-9_]*
        # Or ...
        |
        # Might be a meta variable (usage() will figure it out).
        [A-Z][A-Z0-9_]+
    )
''', re.VERBOSE)

# Compiled regular expression used to recognize options.
OPTION_PATTERN = re.compile(r'^(-\w|--\w+(-\w+)*(=\S+)?)$')

# Initialize a logger for this module.
logger = logging.getLogger(__name__)


def format_usage(usage_text):
    """
    Highlight special items in a usage message.

    :param usage_text: The usage message to process (a string).
    :returns: The usage message with special items highlighted.

    This function highlights the following special items:

    - The initial line of the form "Usage: ..."
    - Short and long command line options
    - Environment variables
    - Meta variables (see :func:`find_meta_variables()`)

    All items are highlighted in the color defined by
    :data:`.HIGHLIGHT_COLOR`.
    """
    # Ugly workaround to avoid circular import errors due to interdependencies
    # between the humanfriendly.terminal and humanfriendly.usage modules.
    from humanfriendly.terminal import ansi_wrap, HIGHLIGHT_COLOR
    formatted_lines = []
    meta_variables = find_meta_variables(usage_text)
    for line in usage_text.strip().splitlines(True):
        if line.startswith(USAGE_MARKER):
            # Highlight the "Usage: ..." line in bold font and color.
            formatted_lines.append(ansi_wrap(line, color=HIGHLIGHT_COLOR))
        else:
            # Highlight options, meta variables and environment variables.
            formatted_lines.append(replace_special_tokens(
                line, meta_variables,
                lambda token: ansi_wrap(token, color=HIGHLIGHT_COLOR),
            ))
    return ''.join(formatted_lines)


def find_meta_variables(usage_text):
    """
    Find the meta variables in the given usage message.

    :param usage_text: The usage message to parse (a string).
    :returns: A list of strings with any meta variables found in the usage
              message.

    When a command line option requires an argument, the convention is to
    format such options as ``--option=ARG``. The text ``ARG`` in this example
    is the meta variable.
    """
    meta_variables = set()
    for match in USAGE_PATTERN.finditer(usage_text):
        token = match.group(0)
        if token.startswith('-'):
            option, _, value = token.partition('=')
            if value:
                meta_variables.add(value)
    return list(meta_variables)


def parse_usage(text):
    """
    Parse a usage message by inferring its structure (and making some assumptions :-).

    :param text: The usage message to parse (a string).
    :returns: A tuple of two lists:

              1. A list of strings with the paragraphs of the usage message's
                 "introduction" (the paragraphs before the documentation of the
                 supported command line options).

              2. A list of strings with pairs of command line options and their
                 descriptions: Item zero is a line listing a supported command
                 line option, item one is the description of that command line
                 option, item two is a line listing another supported command
                 line option, etc.

    Usage messages in general are free format of course, however
    :func:`parse_usage()` assume a certain structure from usage messages in
    order to successfully parse them:

    - The usage message starts with a line ``Usage: ...`` that shows a symbolic
      representation of the way the program is to be invoked.

    - After some free form text a line ``Supported options:`` (surrounded by
      empty lines) precedes the documentation of the supported command line
      options.

    - The command line options are documented as follows::

        -v, --verbose

          Make more noise.

      So all of the variants of the command line option are shown together on a
      separate line, followed by one or more paragraphs describing the option.

    - There are several other minor assumptions, but to be honest I'm not sure if
      anyone other than me is ever going to use this functionality, so for now I
      won't list every intricate detail :-).

      If you're curious anyway, refer to the usage message of the `humanfriendly`
      package (defined in the :mod:`humanfriendly.cli` module) and compare it with
      the usage message you see when you run ``humanfriendly --help`` and the
      generated usage message embedded in the readme.

      Feel free to request more detailed documentation if you're interested in
      using the :mod:`humanfriendly.usage` module outside of the little ecosystem
      of Python packages that I have been building over the past years.
    """
    introduction = []
    documented_options = []
    # Split the raw usage message into paragraphs.
    paragraphs = split_paragraphs(text)
    # Get the paragraphs that are part of the introduction.
    while paragraphs:
        # Check whether we've found the end of the introduction.
        end_of_intro = (paragraphs[0] == START_OF_OPTIONS_MARKER)
        # Append the current paragraph to the introduction.
        introduction.append(paragraphs.pop(0))
        # Stop after we've processed the complete introduction.
        if end_of_intro:
            break
    logger.debug("Parsed introduction: %s", introduction)
    # Parse the paragraphs that document command line options.
    while paragraphs:
        documented_options.append(dedent(paragraphs.pop(0)))
        description = []
        while paragraphs:
            # Check if the next paragraph starts the documentation of another
            # command line option. We split on a comma followed by a space so
            # that our parsing doesn't trip up when the label used for an
            # option's value contains commas.
            tokens = [t.strip() for t in re.split(r',\s', paragraphs[0]) if t and not t.isspace()]
            if all(OPTION_PATTERN.match(t) for t in tokens):
                break
            else:
                description.append(paragraphs.pop(0))
        # Join the description's paragraphs back together so we can remove
        # common leading indentation.
        documented_options.append(dedent('\n\n'.join(description)))
    logger.debug("Parsed options: %s", documented_options)
    return introduction, documented_options


def render_usage(text):
    """
    Reformat a command line program's usage message to reStructuredText_.

    :param text: The plain text usage message (a string).
    :returns: The usage message rendered to reStructuredText_ (a string).
    """
    meta_variables = find_meta_variables(text)
    introduction, options = parse_usage(text)
    output = [render_paragraph(p, meta_variables) for p in introduction]
    if options:
        output.append('\n'.join([
            '.. csv-table::',
            '   :header: Option, Description',
            '   :widths: 30, 70',
            '',
        ]))
        csv_buffer = StringIO()
        csv_writer = csv.writer(csv_buffer)
        while options:
            variants = options.pop(0)
            description = options.pop(0)
            csv_writer.writerow([
                render_paragraph(variants, meta_variables),
                ('\n\n'.join(render_paragraph(p, meta_variables) for p in split_paragraphs(description))).rstrip(),
            ])
        csv_lines = csv_buffer.getvalue().splitlines()
        output.append('\n'.join('   %s' % line for line in csv_lines))
    logger.debug("Rendered output: %s", output)
    return '\n\n'.join(trim_empty_lines(o) for o in output)


def inject_usage(module_name):
    """
    Use cog_ to inject a usage message into a reStructuredText_ file.

    :param module_name: The name of the module whose ``__doc__`` attribute is
                        the source of the usage message (a string).

    This simple wrapper around :func:`render_usage()` makes it very easy to
    inject a reformatted usage message into your documentation using cog_. To
    use it you add a fragment like the following to your ``*.rst`` file::

       .. [[[cog
       .. from humanfriendly.usage import inject_usage
       .. inject_usage('humanfriendly.cli')
       .. ]]]
       .. [[[end]]]

    The lines in the fragment above are single line reStructuredText_ comments
    that are not copied to the output. Their purpose is to instruct cog_ where
    to inject the reformatted usage message. Once you've added these lines to
    your ``*.rst`` file, updating the rendered usage message becomes really
    simple thanks to cog_:

    .. code-block:: sh

       $ cog.py -r README.rst

    This will inject or replace the rendered usage message in your
    ``README.rst`` file with an up to date copy.

    .. _cog: http://nedbatchelder.com/code/cog/
    """
    import cog
    usage_text = import_module(module_name).__doc__
    cog.out("\n" + render_usage(usage_text) + "\n\n")


def render_paragraph(paragraph, meta_variables):
    # Reformat the "Usage:" line to highlight "Usage:" in bold and show the
    # remainder of the line as pre-formatted text.
    if paragraph.startswith(USAGE_MARKER):
        tokens = paragraph.split()
        return "**%s** `%s`" % (tokens[0], ' '.join(tokens[1:]))
    # Reformat the "Supported options:" line to highlight it in bold.
    if paragraph == 'Supported options:':
        return "**%s**" % paragraph
    # Reformat shell transcripts into code blocks.
    if re.match(r'^\s*\$\s+\S', paragraph):
        # Split the paragraph into lines.
        lines = paragraph.splitlines()
        # Check if the paragraph is already indented.
        if not paragraph[0].isspace():
            # If the paragraph isn't already indented we'll indent it now.
            lines = ['  %s' % line for line in lines]
        lines.insert(0, '.. code-block:: sh')
        lines.insert(1, '')
        return "\n".join(lines)
    # The following reformatting applies only to paragraphs which are not
    # indented. Yes this is a hack - for now we assume that indented paragraphs
    # are code blocks, even though this assumption can be wrong.
    if not paragraph[0].isspace():
        # Change UNIX style `quoting' so it doesn't trip up DocUtils.
        paragraph = re.sub("`(.+?)'", r'"\1"', paragraph)
        # Escape asterisks.
        paragraph = paragraph.replace('*', r'\*')
        # Reformat inline tokens.
        paragraph = replace_special_tokens(
            paragraph, meta_variables,
            lambda token: '``%s``' % token,
        )
    return paragraph


def replace_special_tokens(text, meta_variables, replace_fn):
    return USAGE_PATTERN.sub(functools.partial(
        replace_tokens_callback,
        meta_variables=meta_variables,
        replace_fn=replace_fn
    ), text)


def replace_tokens_callback(match, meta_variables, replace_fn):
    token = match.group(0)
    if not (re.match('^[A-Z][A-Z0-9_]+$', token) and token not in meta_variables):
        token = replace_fn(token)
    return token

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\window\numba_.py ===
from __future__ import annotations

import functools
from typing import (
    TYPE_CHECKING,
    Any,
    Callable,
)

import numpy as np

from pandas.compat._optional import import_optional_dependency

from pandas.core.util.numba_ import jit_user_function

if TYPE_CHECKING:
    from pandas._typing import Scalar


@functools.cache
def generate_numba_apply_func(
    func: Callable[..., Scalar],
    nopython: bool,
    nogil: bool,
    parallel: bool,
):
    """
    Generate a numba jitted apply function specified by values from engine_kwargs.

    1. jit the user's function
    2. Return a rolling apply function with the jitted function inline

    Configurations specified in engine_kwargs apply to both the user's
    function _AND_ the rolling apply function.

    Parameters
    ----------
    func : function
        function to be applied to each window and will be JITed
    nopython : bool
        nopython to be passed into numba.jit
    nogil : bool
        nogil to be passed into numba.jit
    parallel : bool
        parallel to be passed into numba.jit

    Returns
    -------
    Numba function
    """
    numba_func = jit_user_function(func)
    if TYPE_CHECKING:
        import numba
    else:
        numba = import_optional_dependency("numba")

    @numba.jit(nopython=nopython, nogil=nogil, parallel=parallel)
    def roll_apply(
        values: np.ndarray,
        begin: np.ndarray,
        end: np.ndarray,
        minimum_periods: int,
        *args: Any,
    ) -> np.ndarray:
        result = np.empty(len(begin))
        for i in numba.prange(len(result)):
            start = begin[i]
            stop = end[i]
            window = values[start:stop]
            count_nan = np.sum(np.isnan(window))
            if len(window) - count_nan >= minimum_periods:
                result[i] = numba_func(window, *args)
            else:
                result[i] = np.nan
        return result

    return roll_apply


@functools.cache
def generate_numba_ewm_func(
    nopython: bool,
    nogil: bool,
    parallel: bool,
    com: float,
    adjust: bool,
    ignore_na: bool,
    deltas: tuple,
    normalize: bool,
):
    """
    Generate a numba jitted ewm mean or sum function specified by values
    from engine_kwargs.

    Parameters
    ----------
    nopython : bool
        nopython to be passed into numba.jit
    nogil : bool
        nogil to be passed into numba.jit
    parallel : bool
        parallel to be passed into numba.jit
    com : float
    adjust : bool
    ignore_na : bool
    deltas : tuple
    normalize : bool

    Returns
    -------
    Numba function
    """
    if TYPE_CHECKING:
        import numba
    else:
        numba = import_optional_dependency("numba")

    @numba.jit(nopython=nopython, nogil=nogil, parallel=parallel)
    def ewm(
        values: np.ndarray,
        begin: np.ndarray,
        end: np.ndarray,
        minimum_periods: int,
    ) -> np.ndarray:
        result = np.empty(len(values))
        alpha = 1.0 / (1.0 + com)
        old_wt_factor = 1.0 - alpha
        new_wt = 1.0 if adjust else alpha

        for i in numba.prange(len(begin)):
            start = begin[i]
            stop = end[i]
            window = values[start:stop]
            sub_result = np.empty(len(window))

            weighted = window[0]
            nobs = int(not np.isnan(weighted))
            sub_result[0] = weighted if nobs >= minimum_periods else np.nan
            old_wt = 1.0

            for j in range(1, len(window)):
                cur = window[j]
                is_observation = not np.isnan(cur)
                nobs += is_observation
                if not np.isnan(weighted):
                    if is_observation or not ignore_na:
                        if normalize:
                            # note that len(deltas) = len(vals) - 1 and deltas[i]
                            # is to be used in conjunction with vals[i+1]
                            old_wt *= old_wt_factor ** deltas[start + j - 1]
                        else:
                            weighted = old_wt_factor * weighted
                        if is_observation:
                            if normalize:
                                # avoid numerical errors on constant series
                                if weighted != cur:
                                    weighted = old_wt * weighted + new_wt * cur
                                    if normalize:
                                        weighted = weighted / (old_wt + new_wt)
                                if adjust:
                                    old_wt += new_wt
                                else:
                                    old_wt = 1.0
                            else:
                                weighted += cur
                elif is_observation:
                    weighted = cur

                sub_result[j] = weighted if nobs >= minimum_periods else np.nan

            result[start:stop] = sub_result

        return result

    return ewm


@functools.cache
def generate_numba_table_func(
    func: Callable[..., np.ndarray],
    nopython: bool,
    nogil: bool,
    parallel: bool,
):
    """
    Generate a numba jitted function to apply window calculations table-wise.

    Func will be passed a M window size x N number of columns array, and
    must return a 1 x N number of columns array. Func is intended to operate
    row-wise, but the result will be transposed for axis=1.

    1. jit the user's function
    2. Return a rolling apply function with the jitted function inline

    Parameters
    ----------
    func : function
        function to be applied to each window and will be JITed
    nopython : bool
        nopython to be passed into numba.jit
    nogil : bool
        nogil to be passed into numba.jit
    parallel : bool
        parallel to be passed into numba.jit

    Returns
    -------
    Numba function
    """
    numba_func = jit_user_function(func)
    if TYPE_CHECKING:
        import numba
    else:
        numba = import_optional_dependency("numba")

    @numba.jit(nopython=nopython, nogil=nogil, parallel=parallel)
    def roll_table(
        values: np.ndarray,
        begin: np.ndarray,
        end: np.ndarray,
        minimum_periods: int,
        *args: Any,
    ):
        result = np.empty((len(begin), values.shape[1]))
        min_periods_mask = np.empty(result.shape)
        for i in numba.prange(len(result)):
            start = begin[i]
            stop = end[i]
            window = values[start:stop]
            count_nan = np.sum(np.isnan(window), axis=0)
            sub_result = numba_func(window, *args)
            nan_mask = len(window) - count_nan >= minimum_periods
            min_periods_mask[i, :] = nan_mask
            result[i, :] = sub_result
        result = np.where(min_periods_mask, result, np.nan)
        return result

    return roll_table


# This function will no longer be needed once numba supports
# axis for all np.nan* agg functions
# https://github.com/numba/numba/issues/1269
@functools.cache
def generate_manual_numpy_nan_agg_with_axis(nan_func):
    if TYPE_CHECKING:
        import numba
    else:
        numba = import_optional_dependency("numba")

    @numba.jit(nopython=True, nogil=True, parallel=True)
    def nan_agg_with_axis(table):
        result = np.empty(table.shape[1])
        for i in numba.prange(table.shape[1]):
            partition = table[:, i]
            result[i] = nan_func(partition)
        return result

    return nan_agg_with_axis


@functools.cache
def generate_numba_ewm_table_func(
    nopython: bool,
    nogil: bool,
    parallel: bool,
    com: float,
    adjust: bool,
    ignore_na: bool,
    deltas: tuple,
    normalize: bool,
):
    """
    Generate a numba jitted ewm mean or sum function applied table wise specified
    by values from engine_kwargs.

    Parameters
    ----------
    nopython : bool
        nopython to be passed into numba.jit
    nogil : bool
        nogil to be passed into numba.jit
    parallel : bool
        parallel to be passed into numba.jit
    com : float
    adjust : bool
    ignore_na : bool
    deltas : tuple
    normalize: bool

    Returns
    -------
    Numba function
    """
    if TYPE_CHECKING:
        import numba
    else:
        numba = import_optional_dependency("numba")

    @numba.jit(nopython=nopython, nogil=nogil, parallel=parallel)
    def ewm_table(
        values: np.ndarray,
        begin: np.ndarray,
        end: np.ndarray,
        minimum_periods: int,
    ) -> np.ndarray:
        alpha = 1.0 / (1.0 + com)
        old_wt_factor = 1.0 - alpha
        new_wt = 1.0 if adjust else alpha
        old_wt = np.ones(values.shape[1])

        result = np.empty(values.shape)
        weighted = values[0].copy()
        nobs = (~np.isnan(weighted)).astype(np.int64)
        result[0] = np.where(nobs >= minimum_periods, weighted, np.nan)
        for i in range(1, len(values)):
            cur = values[i]
            is_observations = ~np.isnan(cur)
            nobs += is_observations.astype(np.int64)
            for j in numba.prange(len(cur)):
                if not np.isnan(weighted[j]):
                    if is_observations[j] or not ignore_na:
                        if normalize:
                            # note that len(deltas) = len(vals) - 1 and deltas[i]
                            # is to be used in conjunction with vals[i+1]
                            old_wt[j] *= old_wt_factor ** deltas[i - 1]
                        else:
                            weighted[j] = old_wt_factor * weighted[j]
                        if is_observations[j]:
                            if normalize:
                                # avoid numerical errors on constant series
                                if weighted[j] != cur[j]:
                                    weighted[j] = (
                                        old_wt[j] * weighted[j] + new_wt * cur[j]
                                    )
                                    if normalize:
                                        weighted[j] = weighted[j] / (old_wt[j] + new_wt)
                                if adjust:
                                    old_wt[j] += new_wt
                                else:
                                    old_wt[j] = 1.0
                            else:
                                weighted[j] += cur[j]
                elif is_observations[j]:
                    weighted[j] = cur[j]

            result[i] = np.where(nobs >= minimum_periods, weighted, np.nan)

        return result

    return ewm_table

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\stats\drv.py ===
from sympy.concrete.summations import (Sum, summation)
from sympy.core.basic import Basic
from sympy.core.cache import cacheit
from sympy.core.function import Lambda
from sympy.core.numbers import I
from sympy.core.relational import (Eq, Ne)
from sympy.core.singleton import S
from sympy.core.symbol import (Dummy, symbols)
from sympy.core.sympify import sympify
from sympy.functions.combinatorial.factorials import factorial
from sympy.functions.elementary.exponential import exp
from sympy.functions.elementary.integers import floor
from sympy.functions.elementary.piecewise import Piecewise
from sympy.logic.boolalg import And
from sympy.polys.polytools import poly
from sympy.series.series import series

from sympy.polys.polyerrors import PolynomialError
from sympy.stats.crv import reduce_rational_inequalities_wrap
from sympy.stats.rv import (NamedArgsMixin, SinglePSpace, SingleDomain,
                            random_symbols, PSpace, ConditionalDomain, RandomDomain,
                            ProductDomain, Distribution)
from sympy.stats.symbolic_probability import Probability
from sympy.sets.fancysets import Range, FiniteSet
from sympy.sets.sets import Union
from sympy.sets.contains import Contains
from sympy.utilities import filldedent
from sympy.core.sympify import _sympify


class DiscreteDistribution(Distribution):
    def __call__(self, *args):
        return self.pdf(*args)


class SingleDiscreteDistribution(DiscreteDistribution, NamedArgsMixin):
    """ Discrete distribution of a single variable.

    Serves as superclass for PoissonDistribution etc....

    Provides methods for pdf, cdf, and sampling

    See Also:
        sympy.stats.crv_types.*
    """

    set = S.Integers

    def __new__(cls, *args):
        args = list(map(sympify, args))
        return Basic.__new__(cls, *args)

    @staticmethod
    def check(*args):
        pass

    @cacheit
    def compute_cdf(self, **kwargs):
        """ Compute the CDF from the PDF.

        Returns a Lambda.
        """
        x = symbols('x', integer=True, cls=Dummy)
        z = symbols('z', real=True, cls=Dummy)
        left_bound = self.set.inf

        # CDF is integral of PDF from left bound to z
        pdf = self.pdf(x)
        cdf = summation(pdf, (x, left_bound, floor(z)), **kwargs)
        # CDF Ensure that CDF left of left_bound is zero
        cdf = Piecewise((cdf, z >= left_bound), (0, True))
        return Lambda(z, cdf)

    def _cdf(self, x):
        return None

    def cdf(self, x, **kwargs):
        """ Cumulative density function """
        if not kwargs:
            cdf = self._cdf(x)
            if cdf is not None:
                return cdf
        return self.compute_cdf(**kwargs)(x)

    @cacheit
    def compute_characteristic_function(self, **kwargs):
        """ Compute the characteristic function from the PDF.

        Returns a Lambda.
        """
        x, t = symbols('x, t', real=True, cls=Dummy)
        pdf = self.pdf(x)
        cf = summation(exp(I*t*x)*pdf, (x, self.set.inf, self.set.sup))
        return Lambda(t, cf)

    def _characteristic_function(self, t):
        return None

    def characteristic_function(self, t, **kwargs):
        """ Characteristic function """
        if not kwargs:
            cf = self._characteristic_function(t)
            if cf is not None:
                return cf
        return self.compute_characteristic_function(**kwargs)(t)

    @cacheit
    def compute_moment_generating_function(self, **kwargs):
        t = Dummy('t', real=True)
        x = Dummy('x', integer=True)
        pdf = self.pdf(x)
        mgf = summation(exp(t*x)*pdf, (x, self.set.inf, self.set.sup))
        return Lambda(t, mgf)

    def _moment_generating_function(self, t):
        return None

    def moment_generating_function(self, t, **kwargs):
        if not kwargs:
            mgf = self._moment_generating_function(t)
            if mgf is not None:
                return mgf
        return self.compute_moment_generating_function(**kwargs)(t)

    @cacheit
    def compute_quantile(self, **kwargs):
        """ Compute the Quantile from the PDF.

        Returns a Lambda.
        """
        x = Dummy('x', integer=True)
        p = Dummy('p', real=True)
        left_bound = self.set.inf
        pdf = self.pdf(x)
        cdf = summation(pdf, (x, left_bound, x), **kwargs)
        set = ((x, p <= cdf), )
        return Lambda(p, Piecewise(*set))

    def _quantile(self, x):
        return None

    def quantile(self, x, **kwargs):
        """ Cumulative density function """
        if not kwargs:
            quantile = self._quantile(x)
            if quantile is not None:
                return quantile
        return self.compute_quantile(**kwargs)(x)

    def expectation(self, expr, var, evaluate=True, **kwargs):
        """ Expectation of expression over distribution """
        # TODO: support discrete sets with non integer stepsizes

        if evaluate:
            try:
                p = poly(expr, var)

                t = Dummy('t', real=True)

                mgf = self.moment_generating_function(t)
                deg = p.degree()
                taylor = poly(series(mgf, t, 0, deg + 1).removeO(), t)
                result = 0
                for k in range(deg+1):
                    result += p.coeff_monomial(var ** k) * taylor.coeff_monomial(t ** k) * factorial(k)

                return result

            except PolynomialError:
                return summation(expr * self.pdf(var),
                                 (var, self.set.inf, self.set.sup), **kwargs)

        else:
            return Sum(expr * self.pdf(var),
                         (var, self.set.inf, self.set.sup), **kwargs)

    def __call__(self, *args):
        return self.pdf(*args)


class DiscreteDomain(RandomDomain):
    """
    A domain with discrete support with step size one.
    Represented using symbols and Range.
    """
    is_Discrete = True

class SingleDiscreteDomain(DiscreteDomain, SingleDomain):
    def as_boolean(self):
        return Contains(self.symbol, self.set)


class ConditionalDiscreteDomain(DiscreteDomain, ConditionalDomain):
    """
    Domain with discrete support of step size one, that is restricted by
    some condition.
    """
    @property
    def set(self):
        rv = self.symbols
        if len(self.symbols) > 1:
            raise NotImplementedError(filldedent('''
                Multivariate conditional domains are not yet implemented.'''))
        rv = list(rv)[0]
        return reduce_rational_inequalities_wrap(self.condition,
            rv).intersect(self.fulldomain.set)


class DiscretePSpace(PSpace):
    is_real = True
    is_Discrete = True

    @property
    def pdf(self):
        return self.density(*self.symbols)

    def where(self, condition):
        rvs = random_symbols(condition)
        assert all(r.symbol in self.symbols for r in rvs)
        if len(rvs) > 1:
            raise NotImplementedError(filldedent('''Multivariate discrete
            random variables are not yet supported.'''))
        conditional_domain = reduce_rational_inequalities_wrap(condition,
            rvs[0])
        conditional_domain = conditional_domain.intersect(self.domain.set)
        return SingleDiscreteDomain(rvs[0].symbol, conditional_domain)

    def probability(self, condition):
        complement = isinstance(condition, Ne)
        if complement:
            condition = Eq(condition.args[0], condition.args[1])
        try:
            _domain = self.where(condition).set
            if condition == False or _domain is S.EmptySet:
                return S.Zero
            if condition == True or _domain == self.domain.set:
                return S.One
            prob = self.eval_prob(_domain)
        except NotImplementedError:
            from sympy.stats.rv import density
            expr = condition.lhs - condition.rhs
            dens = density(expr)
            if not isinstance(dens, DiscreteDistribution):
                from sympy.stats.drv_types import DiscreteDistributionHandmade
                dens = DiscreteDistributionHandmade(dens)
            z = Dummy('z', real=True)
            space = SingleDiscretePSpace(z, dens)
            prob = space.probability(condition.__class__(space.value, 0))
        if prob is None:
            prob = Probability(condition)
        return prob if not complement else S.One - prob

    def eval_prob(self, _domain):
        sym = list(self.symbols)[0]
        if isinstance(_domain, Range):
            n = symbols('n', integer=True)
            inf, sup, step = (r for r in _domain.args)
            summand = ((self.pdf).replace(
              sym, n*step))
            rv = summation(summand,
                (n, inf/step, (sup)/step - 1)).doit()
            return rv
        elif isinstance(_domain, FiniteSet):
            pdf = Lambda(sym, self.pdf)
            rv = sum(pdf(x) for x in _domain)
            return rv
        elif isinstance(_domain, Union):
            rv = sum(self.eval_prob(x) for x in _domain.args)
            return rv

    def conditional_space(self, condition):
        # XXX: Converting from set to tuple. The order matters to Lambda
        # though so we should be starting with a set...
        density = Lambda(tuple(self.symbols), self.pdf/self.probability(condition))
        condition = condition.xreplace({rv: rv.symbol for rv in self.values})
        domain = ConditionalDiscreteDomain(self.domain, condition)
        return DiscretePSpace(domain, density)

class ProductDiscreteDomain(ProductDomain, DiscreteDomain):
    def as_boolean(self):
        return And(*[domain.as_boolean for domain in self.domains])

class SingleDiscretePSpace(DiscretePSpace, SinglePSpace):
    """ Discrete probability space over a single univariate variable """
    is_real = True

    @property
    def set(self):
        return self.distribution.set

    @property
    def domain(self):
        return SingleDiscreteDomain(self.symbol, self.set)

    def sample(self, size=(), library='scipy', seed=None):
        """
        Internal sample method.

        Returns dictionary mapping RandomSymbol to realization value.
        """
        return {self.value: self.distribution.sample(size, library=library, seed=seed)}

    def compute_expectation(self, expr, rvs=None, evaluate=True, **kwargs):
        rvs = rvs or (self.value,)
        if self.value not in rvs:
            return expr

        expr = _sympify(expr)
        expr = expr.xreplace({rv: rv.symbol for rv in rvs})

        x = self.value.symbol
        try:
            return self.distribution.expectation(expr, x, evaluate=evaluate,
                    **kwargs)
        except NotImplementedError:
            return Sum(expr * self.pdf, (x, self.set.inf, self.set.sup),
                    **kwargs)

    def compute_cdf(self, expr, **kwargs):
        if expr == self.value:
            x = Dummy("x", real=True)
            return Lambda(x, self.distribution.cdf(x, **kwargs))
        else:
            raise NotImplementedError()

    def compute_density(self, expr, **kwargs):
        if expr == self.value:
            return self.distribution
        raise NotImplementedError()

    def compute_characteristic_function(self, expr, **kwargs):
        if expr == self.value:
            t = Dummy("t", real=True)
            return Lambda(t, self.distribution.characteristic_function(t, **kwargs))
        else:
            raise NotImplementedError()

    def compute_moment_generating_function(self, expr, **kwargs):
        if expr == self.value:
            t = Dummy("t", real=True)
            return Lambda(t, self.distribution.moment_generating_function(t, **kwargs))
        else:
            raise NotImplementedError()

    def compute_quantile(self, expr, **kwargs):
        if expr == self.value:
            p = Dummy("p", real=True)
            return Lambda(p, self.distribution.quantile(p, **kwargs))
        else:
            raise NotImplementedError()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\werkzeug\datastructures\accept.py ===
from __future__ import annotations

import codecs
import collections.abc as cabc
import re
import typing as t

from .structures import ImmutableList


class Accept(ImmutableList[tuple[str, float]]):
    """An :class:`Accept` object is just a list subclass for lists of
    ``(value, quality)`` tuples.  It is automatically sorted by specificity
    and quality.

    All :class:`Accept` objects work similar to a list but provide extra
    functionality for working with the data.  Containment checks are
    normalized to the rules of that header:

    >>> a = CharsetAccept([('ISO-8859-1', 1), ('utf-8', 0.7)])
    >>> a.best
    'ISO-8859-1'
    >>> 'iso-8859-1' in a
    True
    >>> 'UTF8' in a
    True
    >>> 'utf7' in a
    False

    To get the quality for an item you can use normal item lookup:

    >>> print a['utf-8']
    0.7
    >>> a['utf7']
    0

    .. versionchanged:: 0.5
       :class:`Accept` objects are forced immutable now.

    .. versionchanged:: 1.0.0
       :class:`Accept` internal values are no longer ordered
       alphabetically for equal quality tags. Instead the initial
       order is preserved.

    """

    def __init__(
        self, values: Accept | cabc.Iterable[tuple[str, float]] | None = ()
    ) -> None:
        if values is None:
            super().__init__()
            self.provided = False
        elif isinstance(values, Accept):
            self.provided = values.provided
            super().__init__(values)
        else:
            self.provided = True
            values = sorted(
                values, key=lambda x: (self._specificity(x[0]), x[1]), reverse=True
            )
            super().__init__(values)

    def _specificity(self, value: str) -> tuple[bool, ...]:
        """Returns a tuple describing the value's specificity."""
        return (value != "*",)

    def _value_matches(self, value: str, item: str) -> bool:
        """Check if a value matches a given accept item."""
        return item == "*" or item.lower() == value.lower()

    @t.overload
    def __getitem__(self, key: str) -> float: ...
    @t.overload
    def __getitem__(self, key: t.SupportsIndex) -> tuple[str, float]: ...
    @t.overload
    def __getitem__(self, key: slice) -> list[tuple[str, float]]: ...
    def __getitem__(
        self, key: str | t.SupportsIndex | slice
    ) -> float | tuple[str, float] | list[tuple[str, float]]:
        """Besides index lookup (getting item n) you can also pass it a string
        to get the quality for the item.  If the item is not in the list, the
        returned quality is ``0``.
        """
        if isinstance(key, str):
            return self.quality(key)
        return list.__getitem__(self, key)

    def quality(self, key: str) -> float:
        """Returns the quality of the key.

        .. versionadded:: 0.6
           In previous versions you had to use the item-lookup syntax
           (eg: ``obj[key]`` instead of ``obj.quality(key)``)
        """
        for item, quality in self:
            if self._value_matches(key, item):
                return quality
        return 0

    def __contains__(self, value: str) -> bool:  # type: ignore[override]
        for item, _quality in self:
            if self._value_matches(value, item):
                return True
        return False

    def __repr__(self) -> str:
        pairs_str = ", ".join(f"({x!r}, {y})" for x, y in self)
        return f"{type(self).__name__}([{pairs_str}])"

    def index(self, key: str | tuple[str, float]) -> int:  # type: ignore[override]
        """Get the position of an entry or raise :exc:`ValueError`.

        :param key: The key to be looked up.

        .. versionchanged:: 0.5
           This used to raise :exc:`IndexError`, which was inconsistent
           with the list API.
        """
        if isinstance(key, str):
            for idx, (item, _quality) in enumerate(self):
                if self._value_matches(key, item):
                    return idx
            raise ValueError(key)
        return list.index(self, key)

    def find(self, key: str | tuple[str, float]) -> int:
        """Get the position of an entry or return -1.

        :param key: The key to be looked up.
        """
        try:
            return self.index(key)
        except ValueError:
            return -1

    def values(self) -> cabc.Iterator[str]:
        """Iterate over all values."""
        for item in self:
            yield item[0]

    def to_header(self) -> str:
        """Convert the header set into an HTTP header string."""
        result = []
        for value, quality in self:
            if quality != 1:
                value = f"{value};q={quality}"
            result.append(value)
        return ",".join(result)

    def __str__(self) -> str:
        return self.to_header()

    def _best_single_match(self, match: str) -> tuple[str, float] | None:
        for client_item, quality in self:
            if self._value_matches(match, client_item):
                # self is sorted by specificity descending, we can exit
                return client_item, quality
        return None

    @t.overload
    def best_match(self, matches: cabc.Iterable[str]) -> str | None: ...
    @t.overload
    def best_match(self, matches: cabc.Iterable[str], default: str = ...) -> str: ...
    def best_match(
        self, matches: cabc.Iterable[str], default: str | None = None
    ) -> str | None:
        """Returns the best match from a list of possible matches based
        on the specificity and quality of the client. If two items have the
        same quality and specificity, the one is returned that comes first.

        :param matches: a list of matches to check for
        :param default: the value that is returned if none match
        """
        result = default
        best_quality: float = -1
        best_specificity: tuple[float, ...] = (-1,)
        for server_item in matches:
            match = self._best_single_match(server_item)
            if not match:
                continue
            client_item, quality = match
            specificity = self._specificity(client_item)
            if quality <= 0 or quality < best_quality:
                continue
            # better quality or same quality but more specific => better match
            if quality > best_quality or specificity > best_specificity:
                result = server_item
                best_quality = quality
                best_specificity = specificity
        return result

    @property
    def best(self) -> str | None:
        """The best match as value."""
        if self:
            return self[0][0]

        return None


_mime_split_re = re.compile(r"/|(?:\s*;\s*)")


def _normalize_mime(value: str) -> list[str]:
    return _mime_split_re.split(value.lower())


class MIMEAccept(Accept):
    """Like :class:`Accept` but with special methods and behavior for
    mimetypes.
    """

    def _specificity(self, value: str) -> tuple[bool, ...]:
        return tuple(x != "*" for x in _mime_split_re.split(value))

    def _value_matches(self, value: str, item: str) -> bool:
        # item comes from the client, can't match if it's invalid.
        if "/" not in item:
            return False

        # value comes from the application, tell the developer when it
        # doesn't look valid.
        if "/" not in value:
            raise ValueError(f"invalid mimetype {value!r}")

        # Split the match value into type, subtype, and a sorted list of parameters.
        normalized_value = _normalize_mime(value)
        value_type, value_subtype = normalized_value[:2]
        value_params = sorted(normalized_value[2:])

        # "*/*" is the only valid value that can start with "*".
        if value_type == "*" and value_subtype != "*":
            raise ValueError(f"invalid mimetype {value!r}")

        # Split the accept item into type, subtype, and parameters.
        normalized_item = _normalize_mime(item)
        item_type, item_subtype = normalized_item[:2]
        item_params = sorted(normalized_item[2:])

        # "*/not-*" from the client is invalid, can't match.
        if item_type == "*" and item_subtype != "*":
            return False

        return (
            (item_type == "*" and item_subtype == "*")
            or (value_type == "*" and value_subtype == "*")
        ) or (
            item_type == value_type
            and (
                item_subtype == "*"
                or value_subtype == "*"
                or (item_subtype == value_subtype and item_params == value_params)
            )
        )

    @property
    def accept_html(self) -> bool:
        """True if this object accepts HTML."""
        return "text/html" in self or self.accept_xhtml  # type: ignore[comparison-overlap]

    @property
    def accept_xhtml(self) -> bool:
        """True if this object accepts XHTML."""
        return "application/xhtml+xml" in self or "application/xml" in self  # type: ignore[comparison-overlap]

    @property
    def accept_json(self) -> bool:
        """True if this object accepts JSON."""
        return "application/json" in self  # type: ignore[comparison-overlap]


_locale_delim_re = re.compile(r"[_-]")


def _normalize_lang(value: str) -> list[str]:
    """Process a language tag for matching."""
    return _locale_delim_re.split(value.lower())


class LanguageAccept(Accept):
    """Like :class:`Accept` but with normalization for language tags."""

    def _value_matches(self, value: str, item: str) -> bool:
        return item == "*" or _normalize_lang(value) == _normalize_lang(item)

    @t.overload
    def best_match(self, matches: cabc.Iterable[str]) -> str | None: ...
    @t.overload
    def best_match(self, matches: cabc.Iterable[str], default: str = ...) -> str: ...
    def best_match(
        self, matches: cabc.Iterable[str], default: str | None = None
    ) -> str | None:
        """Given a list of supported values, finds the best match from
        the list of accepted values.

        Language tags are normalized for the purpose of matching, but
        are returned unchanged.

        If no exact match is found, this will fall back to matching
        the first subtag (primary language only), first with the
        accepted values then with the match values. This partial is not
        applied to any other language subtags.

        The default is returned if no exact or fallback match is found.

        :param matches: A list of supported languages to find a match.
        :param default: The value that is returned if none match.
        """
        # Look for an exact match first. If a client accepts "en-US",
        # "en-US" is a valid match at this point.
        result = super().best_match(matches)

        if result is not None:
            return result

        # Fall back to accepting primary tags. If a client accepts
        # "en-US", "en" is a valid match at this point. Need to use
        # re.split to account for 2 or 3 letter codes.
        fallback = Accept(
            [(_locale_delim_re.split(item[0], 1)[0], item[1]) for item in self]
        )
        result = fallback.best_match(matches)

        if result is not None:
            return result

        # Fall back to matching primary tags. If the client accepts
        # "en", "en-US" is a valid match at this point.
        fallback_matches = [_locale_delim_re.split(item, 1)[0] for item in matches]
        result = super().best_match(fallback_matches)

        # Return a value from the original match list. Find the first
        # original value that starts with the matched primary tag.
        if result is not None:
            return next(item for item in matches if item.startswith(result))

        return default


class CharsetAccept(Accept):
    """Like :class:`Accept` but with normalization for charsets."""

    def _value_matches(self, value: str, item: str) -> bool:
        def _normalize(name: str) -> str:
            try:
                return codecs.lookup(name).name
            except LookupError:
                return name.lower()

        return item == "*" or _normalize(value) == _normalize(item)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\reader\excel.py ===
# Copyright (c) 2010-2024 openpyxl


"""Read an xlsx file into Python"""

# Python stdlib imports
from zipfile import ZipFile, ZIP_DEFLATED
from io import BytesIO
import os.path
import warnings

from openpyxl.pivot.table import TableDefinition

# Allow blanket setting of KEEP_VBA for testing
try:
    from ..tests import KEEP_VBA
except ImportError:
    KEEP_VBA = False

# package imports
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.xml.constants import (
    ARC_CORE,
    ARC_CUSTOM,
    ARC_CONTENT_TYPES,
    ARC_WORKBOOK,
    ARC_THEME,
    COMMENTS_NS,
    SHARED_STRINGS,
    XLTM,
    XLTX,
    XLSM,
    XLSX,
)
from openpyxl.cell import MergedCell
from openpyxl.comments.comment_sheet import CommentSheet

from .strings import read_string_table, read_rich_text
from .workbook import WorkbookParser
from openpyxl.styles.stylesheet import apply_stylesheet

from openpyxl.packaging.core import DocumentProperties
from openpyxl.packaging.custom import CustomPropertyList
from openpyxl.packaging.manifest import Manifest, Override

from openpyxl.packaging.relationship import (
    RelationshipList,
    get_dependents,
    get_rels_path,
)

from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet._reader import WorksheetReader
from openpyxl.chartsheet import Chartsheet
from openpyxl.worksheet.table import Table
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing

from openpyxl.xml.functions import fromstring

from .drawings import find_images


SUPPORTED_FORMATS = ('.xlsx', '.xlsm', '.xltx', '.xltm')


def _validate_archive(filename):
    """
    Does a first check whether filename is a string or a file-like
    object. If it is a string representing a filename, a check is done
    for supported formats by checking the given file-extension. If the
    file-extension is not in SUPPORTED_FORMATS an InvalidFileException
    will raised. Otherwise the filename (resp. file-like object) will
    forwarded to zipfile.ZipFile returning a ZipFile-Instance.
    """
    is_file_like = hasattr(filename, 'read')
    if not is_file_like:
        file_format = os.path.splitext(filename)[-1].lower()
        if file_format not in SUPPORTED_FORMATS:
            if file_format == '.xls':
                msg = ('openpyxl does not support the old .xls file format, '
                       'please use xlrd to read this file, or convert it to '
                       'the more recent .xlsx file format.')
            elif file_format == '.xlsb':
                msg = ('openpyxl does not support binary format .xlsb, '
                       'please convert this file to .xlsx format if you want '
                       'to open it with openpyxl')
            else:
                msg = ('openpyxl does not support %s file format, '
                       'please check you can open '
                       'it with Excel first. '
                       'Supported formats are: %s') % (file_format,
                                                       ','.join(SUPPORTED_FORMATS))
            raise InvalidFileException(msg)

    archive = ZipFile(filename, 'r')
    return archive


def _find_workbook_part(package):
    workbook_types = [XLTM, XLTX, XLSM, XLSX]
    for ct in workbook_types:
        part = package.find(ct)
        if part:
            return part

    # some applications reassign the default for application/xml
    defaults = {p.ContentType for p in package.Default}
    workbook_type = defaults & set(workbook_types)
    if workbook_type:
        return Override("/" + ARC_WORKBOOK, workbook_type.pop())

    raise IOError("File contains no valid workbook part")


class ExcelReader:

    """
    Read an Excel package and dispatch the contents to the relevant modules
    """

    def __init__(self, fn, read_only=False, keep_vba=KEEP_VBA,
                 data_only=False, keep_links=True, rich_text=False):
        self.archive = _validate_archive(fn)
        self.valid_files = self.archive.namelist()
        self.read_only = read_only
        self.keep_vba = keep_vba
        self.data_only = data_only
        self.keep_links = keep_links
        self.rich_text = rich_text
        self.shared_strings = []


    def read_manifest(self):
        src = self.archive.read(ARC_CONTENT_TYPES)
        root = fromstring(src)
        self.package = Manifest.from_tree(root)


    def read_strings(self):
        ct = self.package.find(SHARED_STRINGS)
        reader = read_string_table
        if self.rich_text:
            reader = read_rich_text
        if ct is not None:
            strings_path = ct.PartName[1:]
            with self.archive.open(strings_path,) as src:
                self.shared_strings = reader(src)


    def read_workbook(self):
        wb_part = _find_workbook_part(self.package)
        self.parser = WorkbookParser(self.archive, wb_part.PartName[1:], keep_links=self.keep_links)
        self.parser.parse()
        wb = self.parser.wb
        wb._sheets = []
        wb._data_only = self.data_only
        wb._read_only = self.read_only
        wb.template = wb_part.ContentType in (XLTX, XLTM)

        # If are going to preserve the vba then attach a copy of the archive to the
        # workbook so that is available for the save.
        if self.keep_vba:
            wb.vba_archive = ZipFile(BytesIO(), 'a', ZIP_DEFLATED)
            for name in self.valid_files:
                wb.vba_archive.writestr(name, self.archive.read(name))

        if self.read_only:
            wb._archive = self.archive

        self.wb = wb


    def read_properties(self):
        if ARC_CORE in self.valid_files:
            src = fromstring(self.archive.read(ARC_CORE))
            self.wb.properties = DocumentProperties.from_tree(src)


    def read_custom(self):
        if ARC_CUSTOM in self.valid_files:
            src = fromstring(self.archive.read(ARC_CUSTOM))
            self.wb.custom_doc_props = CustomPropertyList.from_tree(src)


    def read_theme(self):
        if ARC_THEME in self.valid_files:
            self.wb.loaded_theme = self.archive.read(ARC_THEME)


    def read_chartsheet(self, sheet, rel):
        sheet_path = rel.target
        rels_path = get_rels_path(sheet_path)
        rels = []
        if rels_path in self.valid_files:
            rels = get_dependents(self.archive, rels_path)

        with self.archive.open(sheet_path, "r") as src:
            xml = src.read()
        node = fromstring(xml)
        cs = Chartsheet.from_tree(node)
        cs._parent = self.wb
        cs.title = sheet.name
        self.wb._add_sheet(cs)

        drawings = rels.find(SpreadsheetDrawing._rel_type)
        for rel in drawings:
            charts, images = find_images(self.archive, rel.target)
            for c in charts:
                cs.add_chart(c)


    def read_worksheets(self):
        comment_warning = """Cell '{0}':{1} is part of a merged range but has a comment which will be removed because merged cells cannot contain any data."""
        for sheet, rel in self.parser.find_sheets():
            if rel.target not in self.valid_files:
                continue

            if "chartsheet" in rel.Type:
                self.read_chartsheet(sheet, rel)
                continue

            rels_path = get_rels_path(rel.target)
            rels = RelationshipList()
            if rels_path in self.valid_files:
                rels = get_dependents(self.archive, rels_path)

            if self.read_only:
                ws = ReadOnlyWorksheet(self.wb, sheet.name, rel.target, self.shared_strings)
                ws.sheet_state = sheet.state
                self.wb._sheets.append(ws)
                continue
            else:
                fh = self.archive.open(rel.target)
                ws = self.wb.create_sheet(sheet.name)
                ws._rels = rels
                ws_parser = WorksheetReader(ws, fh, self.shared_strings, self.data_only, self.rich_text)
                ws_parser.bind_all()
                fh.close()

            # assign any comments to cells
            for r in rels.find(COMMENTS_NS):
                src = self.archive.read(r.target)
                comment_sheet = CommentSheet.from_tree(fromstring(src))
                for ref, comment in comment_sheet.comments:
                    try:
                        ws[ref].comment = comment
                    except AttributeError:
                        c = ws[ref]
                        if isinstance(c, MergedCell):
                            warnings.warn(comment_warning.format(ws.title, c.coordinate))
                            continue

            # preserve link to VML file if VBA
            if self.wb.vba_archive and ws.legacy_drawing:
                ws.legacy_drawing = rels.get(ws.legacy_drawing).target
            else:
                ws.legacy_drawing = None

            for t in ws_parser.tables:
                src = self.archive.read(t)
                xml = fromstring(src)
                table = Table.from_tree(xml)
                ws.add_table(table)

            drawings = rels.find(SpreadsheetDrawing._rel_type)
            for rel in drawings:
                charts, images = find_images(self.archive, rel.target)
                for c in charts:
                    ws.add_chart(c, c.anchor)
                for im in images:
                    ws.add_image(im, im.anchor)

            pivot_rel = rels.find(TableDefinition.rel_type)
            pivot_caches = self.parser.pivot_caches
            for r in pivot_rel:
                pivot_path = r.Target
                src = self.archive.read(pivot_path)
                tree = fromstring(src)
                pivot = TableDefinition.from_tree(tree)
                pivot.cache = pivot_caches[pivot.cacheId]
                ws.add_pivot(pivot)

            ws.sheet_state = sheet.state


    def read(self):
        action = "read manifest"
        try:
            self.read_manifest()
            action = "read strings"
            self.read_strings()
            action = "read workbook"
            self.read_workbook()
            action = "read properties"
            self.read_properties()
            action = "read custom properties"
            self.read_custom()
            action = "read theme"
            self.read_theme()
            action = "read stylesheet"
            apply_stylesheet(self.archive, self.wb)
            action = "read worksheets"
            self.read_worksheets()
            action = "assign names"
            self.parser.assign_names()
            if not self.read_only:
                self.archive.close()
        except ValueError as e:
            raise ValueError(
                f"Unable to read workbook: could not {action} from {self.archive.filename}.\n"
                "This is most probably because the workbook source files contain some invalid XML.\n"
                "Please see the exception for more details."
                ) from e


def load_workbook(filename, read_only=False, keep_vba=KEEP_VBA,
                  data_only=False, keep_links=True, rich_text=False):
    """Open the given filename and return the workbook

    :param filename: the path to open or a file-like object
    :type filename: string or a file-like object open in binary mode c.f., :class:`zipfile.ZipFile`

    :param read_only: optimised for reading, content cannot be edited
    :type read_only: bool

    :param keep_vba: preserve vba content (this does NOT mean you can use it)
    :type keep_vba: bool

    :param data_only: controls whether cells with formulae have either the formula (default) or the value stored the last time Excel read the sheet
    :type data_only: bool

    :param keep_links: whether links to external workbooks should be preserved. The default is True
    :type keep_links: bool

    :param rich_text: if set to True openpyxl will preserve any rich text formatting in cells. The default is False
    :type rich_text: bool

    :rtype: :class:`openpyxl.workbook.Workbook`

    .. note::

        When using lazy load, all worksheets will be :class:`openpyxl.worksheet.iter_worksheet.IterableWorksheet`
        and the returned workbook will be read-only.

    """
    reader = ExcelReader(filename, read_only, keep_vba,
                         data_only, keep_links, rich_text)
    reader.read()
    return reader.wb

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\dtypes\concat.py ===
"""
Utility functions related to concat.
"""
from __future__ import annotations

from typing import (
    TYPE_CHECKING,
    cast,
)
import warnings

import numpy as np

from pandas._libs import lib
from pandas.util._exceptions import find_stack_level

from pandas.core.dtypes.astype import astype_array
from pandas.core.dtypes.cast import (
    common_dtype_categorical_compat,
    find_common_type,
    np_find_common_type,
)
from pandas.core.dtypes.dtypes import CategoricalDtype
from pandas.core.dtypes.generic import (
    ABCCategoricalIndex,
    ABCSeries,
)

if TYPE_CHECKING:
    from collections.abc import Sequence

    from pandas._typing import (
        ArrayLike,
        AxisInt,
        DtypeObj,
    )

    from pandas.core.arrays import (
        Categorical,
        ExtensionArray,
    )


def _is_nonempty(x, axis) -> bool:
    # filter empty arrays
    # 1-d dtypes always are included here
    if x.ndim <= axis:
        return True
    return x.shape[axis] > 0


def concat_compat(
    to_concat: Sequence[ArrayLike], axis: AxisInt = 0, ea_compat_axis: bool = False
) -> ArrayLike:
    """
    provide concatenation of an array of arrays each of which is a single
    'normalized' dtypes (in that for example, if it's object, then it is a
    non-datetimelike and provide a combined dtype for the resulting array that
    preserves the overall dtype if possible)

    Parameters
    ----------
    to_concat : sequence of arrays
    axis : axis to provide concatenation
    ea_compat_axis : bool, default False
        For ExtensionArray compat, behave as if axis == 1 when determining
        whether to drop empty arrays.

    Returns
    -------
    a single array, preserving the combined dtypes
    """
    if len(to_concat) and lib.dtypes_all_equal([obj.dtype for obj in to_concat]):
        # fastpath!
        obj = to_concat[0]
        if isinstance(obj, np.ndarray):
            to_concat_arrs = cast("Sequence[np.ndarray]", to_concat)
            return np.concatenate(to_concat_arrs, axis=axis)

        to_concat_eas = cast("Sequence[ExtensionArray]", to_concat)
        if ea_compat_axis:
            # We have 1D objects, that don't support axis keyword
            return obj._concat_same_type(to_concat_eas)
        elif axis == 0:
            return obj._concat_same_type(to_concat_eas)
        else:
            # e.g. DatetimeArray
            # NB: We are assuming here that ensure_wrapped_if_arraylike has
            #  been called where relevant.
            return obj._concat_same_type(
                # error: Unexpected keyword argument "axis" for "_concat_same_type"
                # of "ExtensionArray"
                to_concat_eas,
                axis=axis,  # type: ignore[call-arg]
            )

    # If all arrays are empty, there's nothing to convert, just short-cut to
    # the concatenation, #3121.
    #
    # Creating an empty array directly is tempting, but the winnings would be
    # marginal given that it would still require shape & dtype calculation and
    # np.concatenate which has them both implemented is compiled.
    orig = to_concat
    non_empties = [x for x in to_concat if _is_nonempty(x, axis)]
    if non_empties and axis == 0 and not ea_compat_axis:
        # ea_compat_axis see GH#39574
        to_concat = non_empties

    any_ea, kinds, target_dtype = _get_result_dtype(to_concat, non_empties)

    if len(to_concat) < len(orig):
        _, _, alt_dtype = _get_result_dtype(orig, non_empties)
        if alt_dtype != target_dtype:
            # GH#39122
            warnings.warn(
                "The behavior of array concatenation with empty entries is "
                "deprecated. In a future version, this will no longer exclude "
                "empty items when determining the result dtype. "
                "To retain the old behavior, exclude the empty entries before "
                "the concat operation.",
                FutureWarning,
                stacklevel=find_stack_level(),
            )

    if target_dtype is not None:
        to_concat = [astype_array(arr, target_dtype, copy=False) for arr in to_concat]

    if not isinstance(to_concat[0], np.ndarray):
        # i.e. isinstance(to_concat[0], ExtensionArray)
        to_concat_eas = cast("Sequence[ExtensionArray]", to_concat)
        cls = type(to_concat[0])
        # GH#53640: eg. for datetime array, axis=1 but 0 is default
        # However, class method `_concat_same_type()` for some classes
        # may not support the `axis` keyword
        if ea_compat_axis or axis == 0:
            return cls._concat_same_type(to_concat_eas)
        else:
            return cls._concat_same_type(
                to_concat_eas,
                axis=axis,  # type: ignore[call-arg]
            )
    else:
        to_concat_arrs = cast("Sequence[np.ndarray]", to_concat)
        result = np.concatenate(to_concat_arrs, axis=axis)

        if not any_ea and "b" in kinds and result.dtype.kind in "iuf":
            # GH#39817 cast to object instead of casting bools to numeric
            result = result.astype(object, copy=False)
    return result


def _get_result_dtype(
    to_concat: Sequence[ArrayLike], non_empties: Sequence[ArrayLike]
) -> tuple[bool, set[str], DtypeObj | None]:
    target_dtype = None

    dtypes = {obj.dtype for obj in to_concat}
    kinds = {obj.dtype.kind for obj in to_concat}

    any_ea = any(not isinstance(x, np.ndarray) for x in to_concat)
    if any_ea:
        # i.e. any ExtensionArrays

        # we ignore axis here, as internally concatting with EAs is always
        # for axis=0
        if len(dtypes) != 1:
            target_dtype = find_common_type([x.dtype for x in to_concat])
            target_dtype = common_dtype_categorical_compat(to_concat, target_dtype)

    elif not len(non_empties):
        # we have all empties, but may need to coerce the result dtype to
        # object if we have non-numeric type operands (numpy would otherwise
        # cast this to float)
        if len(kinds) != 1:
            if not len(kinds - {"i", "u", "f"}) or not len(kinds - {"b", "i", "u"}):
                # let numpy coerce
                pass
            else:
                # coerce to object
                target_dtype = np.dtype(object)
                kinds = {"o"}
    else:
        # error: Argument 1 to "np_find_common_type" has incompatible type
        # "*Set[Union[ExtensionDtype, Any]]"; expected "dtype[Any]"
        target_dtype = np_find_common_type(*dtypes)  # type: ignore[arg-type]

    return any_ea, kinds, target_dtype


def union_categoricals(
    to_union, sort_categories: bool = False, ignore_order: bool = False
) -> Categorical:
    """
    Combine list-like of Categorical-like, unioning categories.

    All categories must have the same dtype.

    Parameters
    ----------
    to_union : list-like
        Categorical, CategoricalIndex, or Series with dtype='category'.
    sort_categories : bool, default False
        If true, resulting categories will be lexsorted, otherwise
        they will be ordered as they appear in the data.
    ignore_order : bool, default False
        If true, the ordered attribute of the Categoricals will be ignored.
        Results in an unordered categorical.

    Returns
    -------
    Categorical

    Raises
    ------
    TypeError
        - all inputs do not have the same dtype
        - all inputs do not have the same ordered property
        - all inputs are ordered and their categories are not identical
        - sort_categories=True and Categoricals are ordered
    ValueError
        Empty list of categoricals passed

    Notes
    -----
    To learn more about categories, see `link
    <https://pandas.pydata.org/pandas-docs/stable/user_guide/categorical.html#unioning>`__

    Examples
    --------
    If you want to combine categoricals that do not necessarily have
    the same categories, `union_categoricals` will combine a list-like
    of categoricals. The new categories will be the union of the
    categories being combined.

    >>> a = pd.Categorical(["b", "c"])
    >>> b = pd.Categorical(["a", "b"])
    >>> pd.api.types.union_categoricals([a, b])
    ['b', 'c', 'a', 'b']
    Categories (3, object): ['b', 'c', 'a']

    By default, the resulting categories will be ordered as they appear
    in the `categories` of the data. If you want the categories to be
    lexsorted, use `sort_categories=True` argument.

    >>> pd.api.types.union_categoricals([a, b], sort_categories=True)
    ['b', 'c', 'a', 'b']
    Categories (3, object): ['a', 'b', 'c']

    `union_categoricals` also works with the case of combining two
    categoricals of the same categories and order information (e.g. what
    you could also `append` for).

    >>> a = pd.Categorical(["a", "b"], ordered=True)
    >>> b = pd.Categorical(["a", "b", "a"], ordered=True)
    >>> pd.api.types.union_categoricals([a, b])
    ['a', 'b', 'a', 'b', 'a']
    Categories (2, object): ['a' < 'b']

    Raises `TypeError` because the categories are ordered and not identical.

    >>> a = pd.Categorical(["a", "b"], ordered=True)
    >>> b = pd.Categorical(["a", "b", "c"], ordered=True)
    >>> pd.api.types.union_categoricals([a, b])
    Traceback (most recent call last):
        ...
    TypeError: to union ordered Categoricals, all categories must be the same

    Ordered categoricals with different categories or orderings can be
    combined by using the `ignore_ordered=True` argument.

    >>> a = pd.Categorical(["a", "b", "c"], ordered=True)
    >>> b = pd.Categorical(["c", "b", "a"], ordered=True)
    >>> pd.api.types.union_categoricals([a, b], ignore_order=True)
    ['a', 'b', 'c', 'c', 'b', 'a']
    Categories (3, object): ['a', 'b', 'c']

    `union_categoricals` also works with a `CategoricalIndex`, or `Series`
    containing categorical data, but note that the resulting array will
    always be a plain `Categorical`

    >>> a = pd.Series(["b", "c"], dtype='category')
    >>> b = pd.Series(["a", "b"], dtype='category')
    >>> pd.api.types.union_categoricals([a, b])
    ['b', 'c', 'a', 'b']
    Categories (3, object): ['b', 'c', 'a']
    """
    from pandas import Categorical
    from pandas.core.arrays.categorical import recode_for_categories

    if len(to_union) == 0:
        raise ValueError("No Categoricals to union")

    def _maybe_unwrap(x):
        if isinstance(x, (ABCCategoricalIndex, ABCSeries)):
            return x._values
        elif isinstance(x, Categorical):
            return x
        else:
            raise TypeError("all components to combine must be Categorical")

    to_union = [_maybe_unwrap(x) for x in to_union]
    first = to_union[0]

    if not lib.dtypes_all_equal([obj.categories.dtype for obj in to_union]):
        raise TypeError("dtype of categories must be the same")

    ordered = False
    if all(first._categories_match_up_to_permutation(other) for other in to_union[1:]):
        # identical categories - fastpath
        categories = first.categories
        ordered = first.ordered

        all_codes = [first._encode_with_my_categories(x)._codes for x in to_union]
        new_codes = np.concatenate(all_codes)

        if sort_categories and not ignore_order and ordered:
            raise TypeError("Cannot use sort_categories=True with ordered Categoricals")

        if sort_categories and not categories.is_monotonic_increasing:
            categories = categories.sort_values()
            indexer = categories.get_indexer(first.categories)

            from pandas.core.algorithms import take_nd

            new_codes = take_nd(indexer, new_codes, fill_value=-1)
    elif ignore_order or all(not c.ordered for c in to_union):
        # different categories - union and recode
        cats = first.categories.append([c.categories for c in to_union[1:]])
        categories = cats.unique()
        if sort_categories:
            categories = categories.sort_values()

        new_codes = [
            recode_for_categories(c.codes, c.categories, categories) for c in to_union
        ]
        new_codes = np.concatenate(new_codes)
    else:
        # ordered - to show a proper error message
        if all(c.ordered for c in to_union):
            msg = "to union ordered Categoricals, all categories must be the same"
            raise TypeError(msg)
        raise TypeError("Categorical.ordered must be the same")

    if ignore_order:
        ordered = False

    dtype = CategoricalDtype(categories=categories, ordered=ordered)
    return Categorical._simple_new(new_codes, dtype=dtype)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\functions\special\bsplines.py ===
from sympy.core import S, sympify
from sympy.core.symbol import (Dummy, symbols)
from sympy.functions import Piecewise, piecewise_fold
from sympy.logic.boolalg import And
from sympy.sets.sets import Interval

from functools import lru_cache


def _ivl(cond, x):
    """return the interval corresponding to the condition

    Conditions in spline's Piecewise give the range over
    which an expression is valid like (lo <= x) & (x <= hi).
    This function returns (lo, hi).
    """
    if isinstance(cond, And) and len(cond.args) == 2:
        a, b = cond.args
        if a.lts == x:
            a, b = b, a
        return a.lts, b.gts
    raise TypeError('unexpected cond type: %s' % cond)


def _add_splines(c, b1, d, b2, x):
    """Construct c*b1 + d*b2."""

    if S.Zero in (b1, c):
        rv = piecewise_fold(d * b2)
    elif S.Zero in (b2, d):
        rv = piecewise_fold(c * b1)
    else:
        new_args = []
        # Just combining the Piecewise without any fancy optimization
        p1 = piecewise_fold(c * b1)
        p2 = piecewise_fold(d * b2)

        # Search all Piecewise arguments except (0, True)
        p2args = list(p2.args[:-1])

        # This merging algorithm assumes the conditions in
        # p1 and p2 are sorted
        for arg in p1.args[:-1]:
            expr = arg.expr
            cond = arg.cond

            lower = _ivl(cond, x)[0]

            # Check p2 for matching conditions that can be merged
            for i, arg2 in enumerate(p2args):
                expr2 = arg2.expr
                cond2 = arg2.cond

                lower_2, upper_2 = _ivl(cond2, x)
                if cond2 == cond:
                    # Conditions match, join expressions
                    expr += expr2
                    # Remove matching element
                    del p2args[i]
                    # No need to check the rest
                    break
                elif lower_2 < lower and upper_2 <= lower:
                    # Check if arg2 condition smaller than arg1,
                    # add to new_args by itself (no match expected
                    # in p1)
                    new_args.append(arg2)
                    del p2args[i]
                    break

            # Checked all, add expr and cond
            new_args.append((expr, cond))

        # Add remaining items from p2args
        new_args.extend(p2args)

        # Add final (0, True)
        new_args.append((0, True))

        rv = Piecewise(*new_args, evaluate=False)

    return rv.expand()


@lru_cache(maxsize=128)
def bspline_basis(d, knots, n, x):
    """
    The $n$-th B-spline at $x$ of degree $d$ with knots.

    Explanation
    ===========

    B-Splines are piecewise polynomials of degree $d$. They are defined on a
    set of knots, which is a sequence of integers or floats.

    Examples
    ========

    The 0th degree splines have a value of 1 on a single interval:

        >>> from sympy import bspline_basis
        >>> from sympy.abc import x
        >>> d = 0
        >>> knots = tuple(range(5))
        >>> bspline_basis(d, knots, 0, x)
        Piecewise((1, (x >= 0) & (x <= 1)), (0, True))

    For a given ``(d, knots)`` there are ``len(knots)-d-1`` B-splines
    defined, that are indexed by ``n`` (starting at 0).

    Here is an example of a cubic B-spline:

        >>> bspline_basis(3, tuple(range(5)), 0, x)
        Piecewise((x**3/6, (x >= 0) & (x <= 1)),
                  (-x**3/2 + 2*x**2 - 2*x + 2/3,
                  (x >= 1) & (x <= 2)),
                  (x**3/2 - 4*x**2 + 10*x - 22/3,
                  (x >= 2) & (x <= 3)),
                  (-x**3/6 + 2*x**2 - 8*x + 32/3,
                  (x >= 3) & (x <= 4)),
                  (0, True))

    By repeating knot points, you can introduce discontinuities in the
    B-splines and their derivatives:

        >>> d = 1
        >>> knots = (0, 0, 2, 3, 4)
        >>> bspline_basis(d, knots, 0, x)
        Piecewise((1 - x/2, (x >= 0) & (x <= 2)), (0, True))

    It is quite time consuming to construct and evaluate B-splines. If
    you need to evaluate a B-spline many times, it is best to lambdify them
    first:

        >>> from sympy import lambdify
        >>> d = 3
        >>> knots = tuple(range(10))
        >>> b0 = bspline_basis(d, knots, 0, x)
        >>> f = lambdify(x, b0)
        >>> y = f(0.5)

    Parameters
    ==========

    d : integer
        degree of bspline

    knots : list of integer values
        list of knots points of bspline

    n : integer
        $n$-th B-spline

    x : symbol

    See Also
    ========

    bspline_basis_set

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/B-spline

    """
    # make sure x has no assumptions so conditions don't evaluate
    xvar = x
    x = Dummy()

    knots = tuple(sympify(k) for k in knots)
    d = int(d)
    n = int(n)
    n_knots = len(knots)
    n_intervals = n_knots - 1
    if n + d + 1 > n_intervals:
        raise ValueError("n + d + 1 must not exceed len(knots) - 1")
    if d == 0:
        result = Piecewise(
            (S.One, Interval(knots[n], knots[n + 1]).contains(x)), (0, True)
        )
    elif d > 0:
        denom = knots[n + d + 1] - knots[n + 1]
        if denom != S.Zero:
            B = (knots[n + d + 1] - x) / denom
            b2 = bspline_basis(d - 1, knots, n + 1, x)
        else:
            b2 = B = S.Zero

        denom = knots[n + d] - knots[n]
        if denom != S.Zero:
            A = (x - knots[n]) / denom
            b1 = bspline_basis(d - 1, knots, n, x)
        else:
            b1 = A = S.Zero

        result = _add_splines(A, b1, B, b2, x)
    else:
        raise ValueError("degree must be non-negative: %r" % n)

    # return result with user-given x
    return result.xreplace({x: xvar})


def bspline_basis_set(d, knots, x):
    """
    Return the ``len(knots)-d-1`` B-splines at *x* of degree *d*
    with *knots*.

    Explanation
    ===========

    This function returns a list of piecewise polynomials that are the
    ``len(knots)-d-1`` B-splines of degree *d* for the given knots.
    This function calls ``bspline_basis(d, knots, n, x)`` for different
    values of *n*.

    Examples
    ========

    >>> from sympy import bspline_basis_set
    >>> from sympy.abc import x
    >>> d = 2
    >>> knots = range(5)
    >>> splines = bspline_basis_set(d, knots, x)
    >>> splines
    [Piecewise((x**2/2, (x >= 0) & (x <= 1)),
               (-x**2 + 3*x - 3/2, (x >= 1) & (x <= 2)),
               (x**2/2 - 3*x + 9/2, (x >= 2) & (x <= 3)),
               (0, True)),
    Piecewise((x**2/2 - x + 1/2, (x >= 1) & (x <= 2)),
              (-x**2 + 5*x - 11/2, (x >= 2) & (x <= 3)),
              (x**2/2 - 4*x + 8, (x >= 3) & (x <= 4)),
              (0, True))]

    Parameters
    ==========

    d : integer
        degree of bspline

    knots : list of integers
        list of knots points of bspline

    x : symbol

    See Also
    ========

    bspline_basis

    """
    n_splines = len(knots) - d - 1
    return [bspline_basis(d, tuple(knots), i, x) for i in range(n_splines)]


def interpolating_spline(d, x, X, Y):
    """
    Return spline of degree *d*, passing through the given *X*
    and *Y* values.

    Explanation
    ===========

    This function returns a piecewise function such that each part is
    a polynomial of degree not greater than *d*. The value of *d*
    must be 1 or greater and the values of *X* must be strictly
    increasing.

    Examples
    ========

    >>> from sympy import interpolating_spline
    >>> from sympy.abc import x
    >>> interpolating_spline(1, x, [1, 2, 4, 7], [3, 6, 5, 7])
    Piecewise((3*x, (x >= 1) & (x <= 2)),
            (7 - x/2, (x >= 2) & (x <= 4)),
            (2*x/3 + 7/3, (x >= 4) & (x <= 7)))
    >>> interpolating_spline(3, x, [-2, 0, 1, 3, 4], [4, 2, 1, 1, 3])
    Piecewise((7*x**3/117 + 7*x**2/117 - 131*x/117 + 2, (x >= -2) & (x <= 1)),
            (10*x**3/117 - 2*x**2/117 - 122*x/117 + 77/39, (x >= 1) & (x <= 4)))

    Parameters
    ==========

    d : integer
        Degree of Bspline strictly greater than equal to one

    x : symbol

    X : list of strictly increasing real values
        list of X coordinates through which the spline passes

    Y : list of real values
        list of corresponding Y coordinates through which the spline passes

    See Also
    ========

    bspline_basis_set, interpolating_poly

    """
    from sympy.solvers.solveset import linsolve
    from sympy.matrices.dense import Matrix

    # Input sanitization
    d = sympify(d)
    if not (d.is_Integer and d.is_positive):
        raise ValueError("Spline degree must be a positive integer, not %s." % d)
    if len(X) != len(Y):
        raise ValueError("Number of X and Y coordinates must be the same.")
    if len(X) < d + 1:
        raise ValueError("Degree must be less than the number of control points.")
    if not all(a < b for a, b in zip(X, X[1:])):
        raise ValueError("The x-coordinates must be strictly increasing.")
    X = [sympify(i) for i in X]

    # Evaluating knots value
    if d.is_odd:
        j = (d + 1) // 2
        interior_knots = X[j:-j]
    else:
        j = d // 2
        interior_knots = [
            (a + b)/2 for a, b in zip(X[j : -j - 1], X[j + 1 : -j])
        ]

    knots = [X[0]] * (d + 1) + list(interior_knots) + [X[-1]] * (d + 1)

    basis = bspline_basis_set(d, knots, x)

    A = [[b.subs(x, v) for b in basis] for v in X]

    coeff = linsolve((Matrix(A), Matrix(Y)), symbols("c0:{}".format(len(X)), cls=Dummy))
    coeff = list(coeff)[0]
    intervals = {c for b in basis for (e, c) in b.args if c != True}

    # Sorting the intervals
    #  ival contains the end-points of each interval
    intervals = sorted(intervals, key=lambda c: _ivl(c, x))

    basis_dicts = [{c: e for (e, c) in b.args} for b in basis]
    spline = []
    for i in intervals:
        piece = sum(
            [c * d.get(i, S.Zero) for (c, d) in zip(coeff, basis_dicts)], S.Zero
        )
        spline.append((piece, i))
    return Piecewise(*spline)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\ntheory\ecm.py ===
from math import log

from sympy.core.random import _randint
from sympy.external.gmpy import gcd, invert, sqrt
from sympy.utilities.misc import as_int
from .generate import sieve, primerange
from .primetest import isprime


#----------------------------------------------------------------------------#
#                                                                            #
#                   Lenstra's Elliptic Curve Factorization                   #
#                                                                            #
#----------------------------------------------------------------------------#


class Point:
    """Montgomery form of Points in an elliptic curve.
    In this form, the addition and doubling of points
    does not need any y-coordinate information thus
    decreasing the number of operations.
    Using Montgomery form we try to perform point addition
    and doubling in least amount of multiplications.

    The elliptic curve used here is of the form
    (E : b*y**2*z = x**3 + a*x**2*z + x*z**2).
    The a_24 parameter is equal to (a + 2)/4.

    References
    ==========

    .. [1] Kris Gaj, Soonhak Kwon, Patrick Baier, Paul Kohlbrenner, Hoang Le, Mohammed Khaleeluddin, Ramakrishna Bachimanchi,
           Implementing the Elliptic Curve Method of Factoring in Reconfigurable Hardware,
           Cryptographic Hardware and Embedded Systems - CHES 2006 (2006), pp. 119-133,
           https://doi.org/10.1007/11894063_10
           https://www.hyperelliptic.org/tanja/SHARCS/talks06/Gaj.pdf

    """

    def __init__(self, x_cord, z_cord, a_24, mod):
        """
        Initial parameters for the Point class.

        Parameters
        ==========

        x_cord : X coordinate of the Point
        z_cord : Z coordinate of the Point
        a_24 : Parameter of the elliptic curve in Montgomery form
        mod : modulus
        """
        self.x_cord = x_cord
        self.z_cord = z_cord
        self.a_24 = a_24
        self.mod = mod

    def __eq__(self, other):
        """Two points are equal if X/Z of both points are equal
        """
        if self.a_24 != other.a_24 or self.mod != other.mod:
            return False
        return self.x_cord * other.z_cord % self.mod ==\
            other.x_cord * self.z_cord % self.mod

    def add(self, Q, diff):
        """
        Add two points self and Q where diff = self - Q. Moreover the assumption
        is self.x_cord*Q.x_cord*(self.x_cord - Q.x_cord) != 0. This algorithm
        requires 6 multiplications. Here the difference between the points
        is already known and using this algorithm speeds up the addition
        by reducing the number of multiplication required. Also in the
        mont_ladder algorithm is constructed in a way so that the difference
        between intermediate points is always equal to the initial point.
        So, we always know what the difference between the point is.


        Parameters
        ==========

        Q : point on the curve in Montgomery form
        diff : self - Q

        Examples
        ========

        >>> from sympy.ntheory.ecm import Point
        >>> p1 = Point(11, 16, 7, 29)
        >>> p2 = Point(13, 10, 7, 29)
        >>> p3 = p2.add(p1, p1)
        >>> p3.x_cord
        23
        >>> p3.z_cord
        17
        """
        u = (self.x_cord - self.z_cord)*(Q.x_cord + Q.z_cord)
        v = (self.x_cord + self.z_cord)*(Q.x_cord - Q.z_cord)
        add, subt = u + v, u - v
        x_cord = diff.z_cord * add * add % self.mod
        z_cord = diff.x_cord * subt * subt % self.mod
        return Point(x_cord, z_cord, self.a_24, self.mod)

    def double(self):
        """
        Doubles a point in an elliptic curve in Montgomery form.
        This algorithm requires 5 multiplications.

        Examples
        ========

        >>> from sympy.ntheory.ecm import Point
        >>> p1 = Point(11, 16, 7, 29)
        >>> p2 = p1.double()
        >>> p2.x_cord
        13
        >>> p2.z_cord
        10
        """
        u = pow(self.x_cord + self.z_cord, 2, self.mod)
        v = pow(self.x_cord - self.z_cord, 2, self.mod)
        diff = u - v
        x_cord = u*v % self.mod
        z_cord = diff*(v + self.a_24*diff) % self.mod
        return Point(x_cord, z_cord, self.a_24, self.mod)

    def mont_ladder(self, k):
        """
        Scalar multiplication of a point in Montgomery form
        using Montgomery Ladder Algorithm.
        A total of 11 multiplications are required in each step of this
        algorithm.

        Parameters
        ==========

        k : The positive integer multiplier

        Examples
        ========

        >>> from sympy.ntheory.ecm import Point
        >>> p1 = Point(11, 16, 7, 29)
        >>> p3 = p1.mont_ladder(3)
        >>> p3.x_cord
        23
        >>> p3.z_cord
        17
        """
        Q = self
        R = self.double()
        for i in bin(k)[3:]:
            if  i  == '1':
                Q = R.add(Q, self)
                R = R.double()
            else:
                R = Q.add(R, self)
                Q = Q.double()
        return Q


def _ecm_one_factor(n, B1=10000, B2=100000, max_curve=200, seed=None):
    """Returns one factor of n using
    Lenstra's 2 Stage Elliptic curve Factorization
    with Suyama's Parameterization. Here Montgomery
    arithmetic is used for fast computation of addition
    and doubling of points in elliptic curve.

    Explanation
    ===========

    This ECM method considers elliptic curves in Montgomery
    form (E : b*y**2*z = x**3 + a*x**2*z + x*z**2) and involves
    elliptic curve operations (mod N), where the elements in
    Z are reduced (mod N). Since N is not a prime, E over FF(N)
    is not really an elliptic curve but we can still do point additions
    and doubling as if FF(N) was a field.

    Stage 1 : The basic algorithm involves taking a random point (P) on an
    elliptic curve in FF(N). The compute k*P using Montgomery ladder algorithm.
    Let q be an unknown factor of N. Then the order of the curve E, |E(FF(q))|,
    might be a smooth number that divides k. Then we have k = l * |E(FF(q))|
    for some l. For any point belonging to the curve E, |E(FF(q))|*P = O,
    hence k*P = l*|E(FF(q))|*P. Thus kP.z_cord = 0 (mod q), and the unknownn
    factor of N (q) can be recovered by taking gcd(kP.z_cord, N).

    Stage 2 : This is a continuation of Stage 1 if k*P != O. The idea utilize
    the fact that even if kP != 0, the value of k might miss just one large
    prime divisor of |E(FF(q))|. In this case we only need to compute the
    scalar multiplication by p to get p*k*P = O. Here a second bound B2
    restrict the size of possible values of p.

    Parameters
    ==========

    n : Number to be Factored. Assume that it is a composite number.
    B1 : Stage 1 Bound. Must be an even number.
    B2 : Stage 2 Bound. Must be an even number.
    max_curve : Maximum number of curves generated

    Returns
    =======

    integer | None : a non-trivial divisor of ``n``. ``None`` if not found

    References
    ==========

    .. [1] Carl Pomerance, Richard Crandall, Prime Numbers: A Computational Perspective,
           2nd Edition (2005), page 344, ISBN:978-0387252827
    """
    randint = _randint(seed)

    # When calculating T, if (B1 - 2*D) is negative, it cannot be calculated.
    D = min(sqrt(B2), B1 // 2 - 1)
    sieve.extend(D)
    beta = [0] * D
    S = [0] * D
    k = 1
    for p in primerange(2, B1 + 1):
        k *= pow(p, int(log(B1, p)))

    # Pre-calculate the prime numbers to be used in stage 2.
    # Using the fact that the x-coordinates of point P and its
    # inverse -P coincide, the number of primes to be checked
    # in stage 2 can be reduced.
    deltas_list = []
    for r in range(B1 + 2*D, B2 + 2*D, 4*D):
        # d in deltas iff r+(2d+1) and/or r-(2d+1) is prime
        deltas = {abs(q - r) >> 1 for q in primerange(r - 2*D, r + 2*D)}
        deltas_list.append(list(deltas))

    for _ in range(max_curve):
        #Suyama's Parametrization
        sigma = randint(6, n - 1)
        u = (sigma**2 - 5) % n
        v = (4*sigma) % n
        u_3 = pow(u, 3, n)

        try:
            # We use the elliptic curve y**2 = x**3 + a*x**2 + x
            # where a = pow(v - u, 3, n)*(3*u + v)*invert(4*u_3*v, n) - 2
            # However, we do not declare a because it is more convenient
            # to use a24 = (a + 2)*invert(4, n) in the calculation.
            a24 = pow(v - u, 3, n)*(3*u + v)*invert(16*u_3*v, n) % n
        except ZeroDivisionError:
            #If the invert(16*u_3*v, n) doesn't exist (i.e., g != 1)
            g = gcd(2*u_3*v, n)
            #If g = n, try another curve
            if g == n:
                continue
            return g

        Q = Point(u_3, pow(v, 3, n), a24, n)
        Q = Q.mont_ladder(k)
        g = gcd(Q.z_cord, n)

        #Stage 1 factor
        if g != 1 and g != n:
            return g
        #Stage 1 failure. Q.z = 0, Try another curve
        elif g == n:
            continue

        #Stage 2 - Improved Standard Continuation
        S[0] = Q
        Q2 = Q.double()
        S[1] = Q2.add(Q, Q)
        beta[0] = (S[0].x_cord*S[0].z_cord) % n
        beta[1] = (S[1].x_cord*S[1].z_cord) % n
        for d in range(2, D):
            S[d] = S[d - 1].add(Q2, S[d - 2])
            beta[d] = (S[d].x_cord*S[d].z_cord) % n
        # i.e., S[i] = Q.mont_ladder(2*i + 1)

        g = 1
        W = Q.mont_ladder(4*D)
        T = Q.mont_ladder(B1 - 2*D)
        R = Q.mont_ladder(B1 + 2*D)
        for deltas in deltas_list:
            # R = Q.mont_ladder(r) where r in range(B1 + 2*D, B2 + 2*D, 4*D)
            alpha = (R.x_cord*R.z_cord) % n
            for delta in deltas:
                # We want to calculate
                # f = R.x_cord * S[delta].z_cord - S[delta].x_cord * R.z_cord
                f = (R.x_cord - S[delta].x_cord)*\
                    (R.z_cord + S[delta].z_cord) - alpha + beta[delta]
                g = (g*f) % n
            T, R = R, R.add(W, T)
        g = gcd(n, g)

        #Stage 2 Factor found
        if g != 1 and g != n:
            return g


def ecm(n, B1=10000, B2=100000, max_curve=200, seed=1234):
    """Performs factorization using Lenstra's Elliptic curve method.

    This function repeatedly calls ``_ecm_one_factor`` to compute the factors
    of n. First all the small factors are taken out using trial division.
    Then ``_ecm_one_factor`` is used to compute one factor at a time.

    Parameters
    ==========

    n : Number to be Factored
    B1 : Stage 1 Bound. Must be an even number.
    B2 : Stage 2 Bound. Must be an even number.
    max_curve : Maximum number of curves generated
    seed : Initialize pseudorandom generator

    Examples
    ========

    >>> from sympy.ntheory import ecm
    >>> ecm(25645121643901801)
    {5394769, 4753701529}
    >>> ecm(9804659461513846513)
    {4641991, 2112166839943}
    """
    from .factor_ import _perfect_power
    n = as_int(n)
    if B1 % 2 != 0 or B2 % 2 != 0:
        raise ValueError("both bounds must be even")
    TF_LIMIT = 100000
    factors = set()
    for prime in sieve.primerange(2, TF_LIMIT):
        if n % prime == 0:
            factors.add(prime)
            while(n % prime == 0):
                n //= prime

    queue = []
    def check(m):
        if isprime(m):
            factors.add(m)
            return
        if result := _perfect_power(m, TF_LIMIT):
            return check(result[0])
        queue.append(m)
    check(n)
    while queue:
        n = queue.pop()
        factor = _ecm_one_factor(n, B1, B2, max_curve, seed)
        if factor is None:
            raise ValueError("Increase the bounds")
        check(factor)
        check(n // factor)
    return factors