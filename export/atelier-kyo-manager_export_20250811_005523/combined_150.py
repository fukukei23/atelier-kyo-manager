
# === atelier-kyo-manager/.venv_backup\Lib\site-packages\numpy\f2py\_backends\_meson.py ===
import errno
import os
import re
import shutil
import subprocess
import sys
from itertools import chain
from pathlib import Path
from string import Template

from ._backend import Backend


class MesonTemplate:
    """Template meson build file generation class."""

    def __init__(
        self,
        modulename: str,
        sources: list[Path],
        deps: list[str],
        libraries: list[str],
        library_dirs: list[Path],
        include_dirs: list[Path],
        object_files: list[Path],
        linker_args: list[str],
        fortran_args: list[str],
        build_type: str,
        python_exe: str,
    ):
        self.modulename = modulename
        self.build_template_path = (
            Path(__file__).parent.absolute() / "meson.build.template"
        )
        self.sources = sources
        self.deps = deps
        self.libraries = libraries
        self.library_dirs = library_dirs
        if include_dirs is not None:
            self.include_dirs = include_dirs
        else:
            self.include_dirs = []
        self.substitutions = {}
        self.objects = object_files
        # Convert args to '' wrapped variant for meson
        self.fortran_args = [
            f"'{x}'" if not (x.startswith("'") and x.endswith("'")) else x
            for x in fortran_args
        ]
        self.pipeline = [
            self.initialize_template,
            self.sources_substitution,
            self.deps_substitution,
            self.include_substitution,
            self.libraries_substitution,
            self.fortran_args_substitution,
        ]
        self.build_type = build_type
        self.python_exe = python_exe
        self.indent = " " * 21

    def meson_build_template(self) -> str:
        if not self.build_template_path.is_file():
            raise FileNotFoundError(
                errno.ENOENT,
                "Meson build template"
                f" {self.build_template_path.absolute()}"
                " does not exist.",
            )
        return self.build_template_path.read_text()

    def initialize_template(self) -> None:
        self.substitutions["modulename"] = self.modulename
        self.substitutions["buildtype"] = self.build_type
        self.substitutions["python"] = self.python_exe

    def sources_substitution(self) -> None:
        self.substitutions["source_list"] = ",\n".join(
            [f"{self.indent}'''{source}'''," for source in self.sources]
        )

    def deps_substitution(self) -> None:
        self.substitutions["dep_list"] = f",\n{self.indent}".join(
            [f"{self.indent}dependency('{dep}')," for dep in self.deps]
        )

    def libraries_substitution(self) -> None:
        self.substitutions["lib_dir_declarations"] = "\n".join(
            [
                f"lib_dir_{i} = declare_dependency(link_args : ['''-L{lib_dir}'''])"
                for i, lib_dir in enumerate(self.library_dirs)
            ]
        )

        self.substitutions["lib_declarations"] = "\n".join(
            [
                f"{lib.replace('.', '_')} = declare_dependency(link_args : ['-l{lib}'])"
                for lib in self.libraries
            ]
        )

        self.substitutions["lib_list"] = f"\n{self.indent}".join(
            [f"{self.indent}{lib.replace('.', '_')}," for lib in self.libraries]
        )
        self.substitutions["lib_dir_list"] = f"\n{self.indent}".join(
            [f"{self.indent}lib_dir_{i}," for i in range(len(self.library_dirs))]
        )

    def include_substitution(self) -> None:
        self.substitutions["inc_list"] = f",\n{self.indent}".join(
            [f"{self.indent}'''{inc}'''," for inc in self.include_dirs]
        )

    def fortran_args_substitution(self) -> None:
        if self.fortran_args:
            self.substitutions["fortran_args"] = (
                f"{self.indent}fortran_args: [{', '.join(list(self.fortran_args))}],"
            )
        else:
            self.substitutions["fortran_args"] = ""

    def generate_meson_build(self):
        for node in self.pipeline:
            node()
        template = Template(self.meson_build_template())
        meson_build = template.substitute(self.substitutions)
        meson_build = meson_build.replace(",,", ",")
        return meson_build


class MesonBackend(Backend):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.dependencies = self.extra_dat.get("dependencies", [])
        self.meson_build_dir = "bbdir"
        self.build_type = (
            "debug" if any("debug" in flag for flag in self.fc_flags) else "release"
        )
        self.fc_flags = _get_flags(self.fc_flags)

    def _move_exec_to_root(self, build_dir: Path):
        walk_dir = Path(build_dir) / self.meson_build_dir
        path_objects = chain(
            walk_dir.glob(f"{self.modulename}*.so"),
            walk_dir.glob(f"{self.modulename}*.pyd"),
            walk_dir.glob(f"{self.modulename}*.dll"),
        )
        # Same behavior as distutils
        # https://github.com/numpy/numpy/issues/24874#issuecomment-1835632293
        for path_object in path_objects:
            dest_path = Path.cwd() / path_object.name
            if dest_path.exists():
                dest_path.unlink()
            shutil.copy2(path_object, dest_path)
            os.remove(path_object)

    def write_meson_build(self, build_dir: Path) -> None:
        """Writes the meson build file at specified location"""
        meson_template = MesonTemplate(
            self.modulename,
            self.sources,
            self.dependencies,
            self.libraries,
            self.library_dirs,
            self.include_dirs,
            self.extra_objects,
            self.flib_flags,
            self.fc_flags,
            self.build_type,
            sys.executable,
        )
        src = meson_template.generate_meson_build()
        Path(build_dir).mkdir(parents=True, exist_ok=True)
        meson_build_file = Path(build_dir) / "meson.build"
        meson_build_file.write_text(src)
        return meson_build_file

    def _run_subprocess_command(self, command, cwd):
        subprocess.run(command, cwd=cwd, check=True)

    def run_meson(self, build_dir: Path):
        setup_command = ["meson", "setup", self.meson_build_dir]
        self._run_subprocess_command(setup_command, build_dir)
        compile_command = ["meson", "compile", "-C", self.meson_build_dir]
        self._run_subprocess_command(compile_command, build_dir)

    def compile(self) -> None:
        self.sources = _prepare_sources(self.modulename, self.sources, self.build_dir)
        self.write_meson_build(self.build_dir)
        self.run_meson(self.build_dir)
        self._move_exec_to_root(self.build_dir)


def _prepare_sources(mname, sources, bdir):
    extended_sources = sources.copy()
    Path(bdir).mkdir(parents=True, exist_ok=True)
    # Copy sources
    for source in sources:
        if Path(source).exists() and Path(source).is_file():
            shutil.copy(source, bdir)
    generated_sources = [
        Path(f"{mname}module.c"),
        Path(f"{mname}-f2pywrappers2.f90"),
        Path(f"{mname}-f2pywrappers.f"),
    ]
    bdir = Path(bdir)
    for generated_source in generated_sources:
        if generated_source.exists():
            shutil.copy(generated_source, bdir / generated_source.name)
            extended_sources.append(generated_source.name)
            generated_source.unlink()
    extended_sources = [
        Path(source).name
        for source in extended_sources
        if not Path(source).suffix == ".pyf"
    ]
    return extended_sources


def _get_flags(fc_flags):
    flag_values = []
    flag_pattern = re.compile(r"--f(77|90)flags=(.*)")
    for flag in fc_flags:
        match_result = flag_pattern.match(flag)
        if match_result:
            values = match_result.group(2).strip().split()
            values = [val.strip("'\"") for val in values]
            flag_values.extend(values)
    # Hacky way to preserve order of flags
    unique_flags = list(dict.fromkeys(flag_values))
    return unique_flags

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\quantization\operators\matmul.py ===
import itertools
import logging

import onnx
from onnx import onnx_pb as onnx_proto

from ..quant_utils import TENSOR_NAME_QUANT_SUFFIX, QuantizedValue, QuantizedValueType, find_by_name, get_mul_node
from .base_operator import QuantOperatorBase
from .qdq_base_operator import QDQOperatorBase


class QOpMatMul(QuantOperatorBase):
    def __init__(self, onnx_quantizer, onnx_node):
        super().__init__(onnx_quantizer, onnx_node)

    def should_quantize(self):
        if not self.quantizer.should_quantize_node(self.node):
            logging.debug(f"Ignore MatMul {self.node.name}]")
            return False

        if (not self.quantizer.is_float_tensor(self.node.input[1])) and (
            not self.quantizer.is_float_tensor(self.node.input[0])
        ):
            logging.info(f"Ignore MatMul due to non float inputs {self.node.name}]")
            return False

        # do not quantize non-constant B matrices for matmul
        if self.quantizer.q_matmul_const_b_only:
            if not self.quantizer.find_initializer_in_path(self.node.input[1]):
                logging.info(f"Ignore MatMul due to non constant B: {self.quantizer.graph_scope}[{self.node.name}]")
                return False
        return True


"""
    Used when quantize mode is QuantizationMode.IntegerOps.
"""


class MatMulInteger(QOpMatMul):
    def __init__(self, onnx_quantizer, onnx_node):
        super().__init__(onnx_quantizer, onnx_node)

    def quantize(self):
        node = self.node
        assert node.op_type == "MatMul"
        #  Get Quantized from both activation(input[0]) and weight(input[1])
        (
            quantized_input_names,
            zero_point_names,
            scale_names,
            nodes,
        ) = self.quantizer.quantize_activation(node, [0])

        (
            quantized_input_names_weight,
            zero_point_names_weight,
            scale_names_weight,
            nodes_weight,
        ) = self.quantizer.quantize_weight(node, [1], reduce_range=True, op_level_per_channel=True)
        quantized_input_names.extend(quantized_input_names_weight)
        zero_point_names.extend(zero_point_names_weight)
        scale_names.extend(scale_names_weight)
        nodes.extend(nodes_weight)

        matmul_integer_output = node.output[0] + "_output_quantized"
        matmul_integer_name = node.name + "_quant" if node.name else ""
        matmul_integer_node = onnx.helper.make_node(
            "MatMulInteger",
            quantized_input_names + zero_point_names,
            [matmul_integer_output],
            matmul_integer_name,
        )
        nodes.append(matmul_integer_node)

        # Add cast operation to cast matmulInteger output to float.
        cast_op_output = matmul_integer_output + "_cast_output"
        otype = self.quantizer.get_tensor_type(node.output[0], mandatory=True)
        cast_node = onnx.helper.make_node(
            "Cast",
            [matmul_integer_output],
            [cast_op_output],
            matmul_integer_output + "_cast",
            to=otype,
        )
        nodes.append(cast_node)

        # Add mul operation to multiply scales of two inputs.
        assert len(scale_names) == 2
        scales_mul_op = (
            matmul_integer_name + "_scales_mul"
            if matmul_integer_name
            else scale_names[0] + "_" + scale_names[1] + "_mul"
        )

        scales_mul_node = find_by_name(scales_mul_op, self.quantizer.new_nodes)
        if scales_mul_node is None:
            scales_mul_node = get_mul_node(scale_names, scales_mul_op + ":0", scales_mul_op)
            nodes.append(scales_mul_node)

        scales_mul_op_output = scales_mul_node.output[0]

        # Add mul operation to multiply mul_scales_op result with output of MatMulInteger
        # and make the output of this node the same as output of original matmul node.
        output_scale_mul_op = ""
        if matmul_integer_name:
            output_scale_mul_op = matmul_integer_name + "_output_scale_mul"
        nodes.append(
            get_mul_node(
                [cast_op_output, scales_mul_op_output],
                node.output[0],
                output_scale_mul_op,
            )
        )
        self.quantizer.new_nodes += nodes


"""
    Used when quantize mode is QuantizationMode.QLinearOps
"""


class QLinearMatMul(QOpMatMul):
    def __init__(self, onnx_quantizer, onnx_node):
        super().__init__(onnx_quantizer, onnx_node)

    def quantize(self):
        node = self.node
        assert node.op_type == "MatMul"
        #  Get Quantized from both activation(input[0]) and weight(input[1])
        (
            quantized_input_names,
            zero_point_names,
            scale_names,
            nodes,
        ) = self.quantizer.quantize_activation(node, [0])

        (
            quantized_input_names_weight,
            zero_point_names_weight,
            scale_names_weight,
            nodes_weight,
        ) = self.quantizer.quantize_weight(node, [1], reduce_range=True, op_level_per_channel=True)
        quantized_input_names.extend(quantized_input_names_weight)
        zero_point_names.extend(zero_point_names_weight)
        scale_names.extend(scale_names_weight)

        nodes.extend(nodes_weight)
        (
            data_found,
            output_scale_name,
            output_zp_name,
            _,
            _,
        ) = self.quantizer._get_quantization_params(node.output[0])
        if not data_found or quantized_input_names is None:
            return super().quantize()

        qlinear_matmul_output = node.output[0] + TENSOR_NAME_QUANT_SUFFIX
        qlinear_matmul_name = node.name + "_quant" if node.name else ""

        qlinear_matmul_inputs = []
        # Input 0
        qlinear_matmul_inputs.append(quantized_input_names[0])
        qlinear_matmul_inputs.append(scale_names[0])
        qlinear_matmul_inputs.append(zero_point_names[0])
        # Input 1
        qlinear_matmul_inputs.append(quantized_input_names[1])
        qlinear_matmul_inputs.append(scale_names[1])
        qlinear_matmul_inputs.append(zero_point_names[1])
        # Output quantization parameter
        qlinear_matmul_inputs.append(output_scale_name)
        qlinear_matmul_inputs.append(output_zp_name)

        domain = (
            "com.microsoft"
            if self.quantizer.weight_qType
            in {
                onnx_proto.TensorProto.FLOAT8E4M3FN,
                onnx_proto.TensorProto.FLOAT8E4M3FNUZ,
                onnx_proto.TensorProto.FLOAT8E5M2,
                onnx_proto.TensorProto.FLOAT8E5M2FNUZ,
            }
            else ""
        )
        qlinear_matmul_node = onnx.helper.make_node(
            "QLinearMatMul",
            qlinear_matmul_inputs,
            [qlinear_matmul_output],
            qlinear_matmul_name,
            domain=domain,
        )
        nodes.append(qlinear_matmul_node)

        # Create an entry for this quantized value
        q_output = QuantizedValue(
            node.output[0],
            qlinear_matmul_output,
            output_scale_name,
            output_zp_name,
            QuantizedValueType.Input,
        )
        self.quantizer.quantized_value_map[node.output[0]] = q_output

        self.quantizer.new_nodes += nodes


class QDQMatMul(QDQOperatorBase):
    def __init__(self, onnx_quantizer, onnx_node):
        super().__init__(onnx_quantizer, onnx_node)

    def quantize(self):
        node = self.node
        assert node.op_type == "MatMul"

        if self.disable_qdq_for_node_output:
            nodes_to_iterate = node.input
        else:
            nodes_to_iterate = itertools.chain(node.input, node.output)

        for tensor_name in nodes_to_iterate:
            if find_by_name(tensor_name, self.quantizer.model.initializer()):
                is_per_channel, channel_axis = self.quantizer.is_tensor_per_channel(
                    tensor_name, default_axis=1, op_type=node.op_type
                )
                if is_per_channel:
                    self.quantizer.quantize_weight_tensor_per_channel(tensor_name, channel_axis)
                else:
                    self.quantizer.quantize_weight_tensor(tensor_name)
            else:
                self.quantizer.quantize_activation_tensor(tensor_name)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\paulialgebra.py ===
"""
This module implements Pauli algebra by subclassing Symbol. Only algebraic
properties of Pauli matrices are used (we do not use the Matrix class).

See the documentation to the class Pauli for examples.

References
==========

.. [1] https://en.wikipedia.org/wiki/Pauli_matrices
"""

from sympy.core.add import Add
from sympy.core.mul import Mul
from sympy.core.numbers import I
from sympy.core.power import Pow
from sympy.core.symbol import Symbol
from sympy.physics.quantum import TensorProduct

__all__ = ['evaluate_pauli_product']


def delta(i, j):
    """
    Returns 1 if ``i == j``, else 0.

    This is used in the multiplication of Pauli matrices.

    Examples
    ========

    >>> from sympy.physics.paulialgebra import delta
    >>> delta(1, 1)
    1
    >>> delta(2, 3)
    0
    """
    if i == j:
        return 1
    else:
        return 0


def epsilon(i, j, k):
    """
    Return 1 if i,j,k is equal to (1,2,3), (2,3,1), or (3,1,2);
    -1 if ``i``,``j``,``k`` is equal to (1,3,2), (3,2,1), or (2,1,3);
    else return 0.

    This is used in the multiplication of Pauli matrices.

    Examples
    ========

    >>> from sympy.physics.paulialgebra import epsilon
    >>> epsilon(1, 2, 3)
    1
    >>> epsilon(1, 3, 2)
    -1
    """
    if (i, j, k) in ((1, 2, 3), (2, 3, 1), (3, 1, 2)):
        return 1
    elif (i, j, k) in ((1, 3, 2), (3, 2, 1), (2, 1, 3)):
        return -1
    else:
        return 0


class Pauli(Symbol):
    """
    The class representing algebraic properties of Pauli matrices.

    Explanation
    ===========

    The symbol used to display the Pauli matrices can be changed with an
    optional parameter ``label="sigma"``. Pauli matrices with different
    ``label`` attributes cannot multiply together.

    If the left multiplication of symbol or number with Pauli matrix is needed,
    please use parentheses  to separate Pauli and symbolic multiplication
    (for example: 2*I*(Pauli(3)*Pauli(2))).

    Another variant is to use evaluate_pauli_product function to evaluate
    the product of Pauli matrices and other symbols (with commutative
    multiply rules).

    See Also
    ========

    evaluate_pauli_product

    Examples
    ========

    >>> from sympy.physics.paulialgebra import Pauli
    >>> Pauli(1)
    sigma1
    >>> Pauli(1)*Pauli(2)
    I*sigma3
    >>> Pauli(1)*Pauli(1)
    1
    >>> Pauli(3)**4
    1
    >>> Pauli(1)*Pauli(2)*Pauli(3)
    I

    >>> from sympy.physics.paulialgebra import Pauli
    >>> Pauli(1, label="tau")
    tau1
    >>> Pauli(1)*Pauli(2, label="tau")
    sigma1*tau2
    >>> Pauli(1, label="tau")*Pauli(2, label="tau")
    I*tau3

    >>> from sympy import I
    >>> I*(Pauli(2)*Pauli(3))
    -sigma1

    >>> from sympy.physics.paulialgebra import evaluate_pauli_product
    >>> f = I*Pauli(2)*Pauli(3)
    >>> f
    I*sigma2*sigma3
    >>> evaluate_pauli_product(f)
    -sigma1
    """

    __slots__ = ("i", "label")

    def __new__(cls, i, label="sigma"):
        if i not in [1, 2, 3]:
            raise IndexError("Invalid Pauli index")
        obj = Symbol.__new__(cls, "%s%d" %(label,i), commutative=False, hermitian=True)
        obj.i = i
        obj.label = label
        return obj

    def __getnewargs_ex__(self):
        return (self.i, self.label), {}

    def _hashable_content(self):
        return (self.i, self.label)

    # FIXME don't work for -I*Pauli(2)*Pauli(3)
    def __mul__(self, other):
        if isinstance(other, Pauli):
            j = self.i
            k = other.i
            jlab = self.label
            klab = other.label

            if jlab == klab:
                return delta(j, k) \
                    + I*epsilon(j, k, 1)*Pauli(1,jlab) \
                    + I*epsilon(j, k, 2)*Pauli(2,jlab) \
                    + I*epsilon(j, k, 3)*Pauli(3,jlab)
        return super().__mul__(other)

    def _eval_power(b, e):
        if e.is_Integer and e.is_positive:
            return super().__pow__(int(e) % 2)


def evaluate_pauli_product(arg):
    '''Help function to evaluate Pauli matrices product
    with symbolic objects.

    Parameters
    ==========

    arg: symbolic expression that contains Paulimatrices

    Examples
    ========

    >>> from sympy.physics.paulialgebra import Pauli, evaluate_pauli_product
    >>> from sympy import I
    >>> evaluate_pauli_product(I*Pauli(1)*Pauli(2))
    -sigma3

    >>> from sympy.abc import x
    >>> evaluate_pauli_product(x**2*Pauli(2)*Pauli(1))
    -I*x**2*sigma3
    '''
    start = arg
    end = arg

    if isinstance(arg, Pow) and isinstance(arg.args[0], Pauli):
        if arg.args[1].is_odd:
            return arg.args[0]
        else:
            return 1

    if isinstance(arg, Add):
        return Add(*[evaluate_pauli_product(part) for part in arg.args])

    if isinstance(arg, TensorProduct):
        return TensorProduct(*[evaluate_pauli_product(part) for part in arg.args])

    elif not(isinstance(arg, Mul)):
        return arg

    while not start == end or start == arg and end == arg:
        start = end

        tmp = start.as_coeff_mul()
        sigma_product = 1
        com_product = 1
        keeper = 1

        for el in tmp[1]:
            if isinstance(el, Pauli):
                sigma_product *= el
            elif not el.is_commutative:
                if isinstance(el, Pow) and isinstance(el.args[0], Pauli):
                    if el.args[1].is_odd:
                        sigma_product *= el.args[0]
                elif isinstance(el, TensorProduct):
                    keeper = keeper*sigma_product*\
                        TensorProduct(
                            *[evaluate_pauli_product(part) for part in el.args]
                        )
                    sigma_product = 1
                else:
                    keeper = keeper*sigma_product*el
                    sigma_product = 1
            else:
                com_product *= el
        end = tmp[0]*keeper*sigma_product*com_product
        if end == arg: break
    return end

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\chromium\webdriver.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

from typing import Optional

from selenium.webdriver.chromium.remote_connection import ChromiumRemoteConnection
from selenium.webdriver.common.driver_finder import DriverFinder
from selenium.webdriver.common.options import ArgOptions
from selenium.webdriver.common.service import Service
from selenium.webdriver.remote.command import Command
from selenium.webdriver.remote.webdriver import WebDriver as RemoteWebDriver


class ChromiumDriver(RemoteWebDriver):
    """Controls the WebDriver instance of ChromiumDriver and allows you to
    drive the browser."""

    def __init__(
        self,
        browser_name: Optional[str] = None,
        vendor_prefix: Optional[str] = None,
        options: ArgOptions = ArgOptions(),
        service: Optional[Service] = None,
        keep_alive: bool = True,
    ) -> None:
        """Creates a new WebDriver instance of the ChromiumDriver. Starts the
        service and then creates new WebDriver instance of ChromiumDriver.

        :Args:
         - browser_name - Browser name used when matching capabilities.
         - vendor_prefix - Company prefix to apply to vendor-specific WebDriver extension commands.
         - options - this takes an instance of ChromiumOptions
         - service - Service object for handling the browser driver if you need to pass extra details
         - keep_alive - Whether to configure ChromiumRemoteConnection to use HTTP keep-alive.
        """
        self.service = service

        finder = DriverFinder(self.service, options)
        if finder.get_browser_path():
            options.binary_location = finder.get_browser_path()
            options.browser_version = None

        self.service.path = self.service.env_path() or finder.get_driver_path()
        self.service.start()

        executor = ChromiumRemoteConnection(
            remote_server_addr=self.service.service_url,
            browser_name=browser_name,
            vendor_prefix=vendor_prefix,
            keep_alive=keep_alive,
            ignore_proxy=options._ignore_local_proxy,
        )

        try:
            super().__init__(command_executor=executor, options=options)
        except Exception:
            self.quit()
            raise

        self._is_remote = False

    def launch_app(self, id):
        """Launches Chromium app specified by id."""
        return self.execute("launchApp", {"id": id})

    def get_network_conditions(self):
        """Gets Chromium network emulation settings.

        :Returns:
            A dict.
            For example:     {'latency': 4, 'download_throughput': 2, 'upload_throughput': 2, 'offline': False}
        """
        return self.execute("getNetworkConditions")["value"]

    def set_network_conditions(self, **network_conditions) -> None:
        """Sets Chromium network emulation settings.

        :Args:
         - network_conditions: A dict with conditions specification.

        :Usage:
            ::

                driver.set_network_conditions(
                    offline=False,
                    latency=5,  # additional latency (ms)
                    download_throughput=500 * 1024,  # maximal throughput
                    upload_throughput=500 * 1024,
                )  # maximal throughput

            Note: 'throughput' can be used to set both (for download and upload).
        """
        self.execute("setNetworkConditions", {"network_conditions": network_conditions})

    def delete_network_conditions(self) -> None:
        """Resets Chromium network emulation settings."""
        self.execute("deleteNetworkConditions")

    def set_permissions(self, name: str, value: str) -> None:
        """Sets Applicable Permission.

        :Args:
         - name: The item to set the permission on.
         - value: The value to set on the item

        :Usage:
            ::

                driver.set_permissions("clipboard-read", "denied")
        """
        self.execute("setPermissions", {"descriptor": {"name": name}, "state": value})

    def execute_cdp_cmd(self, cmd: str, cmd_args: dict):
        """Execute Chrome Devtools Protocol command and get returned result The
        command and command args should follow chrome devtools protocol
        domains/commands, refer to link
        https://chromedevtools.github.io/devtools-protocol/

        :Args:
         - cmd: A str, command name
         - cmd_args: A dict, command args. empty dict {} if there is no command args
        :Usage:
            ::

                driver.execute_cdp_cmd('Network.getResponseBody', {'requestId': requestId})
        :Returns:
            A dict, empty dict {} if there is no result to return.
            For example to getResponseBody:
            {'base64Encoded': False, 'body': 'response body string'}
        """
        return super().execute_cdp_cmd(cmd, cmd_args)

    def get_sinks(self) -> list:
        """:Returns: A list of sinks available for Cast."""
        return self.execute("getSinks")["value"]

    def get_issue_message(self):
        """:Returns: An error message when there is any issue in a Cast
        session."""
        return self.execute("getIssueMessage")["value"]

    @property
    def log_types(self):
        """Gets a list of the available log types.

        Example:
        --------
        >>> driver.log_types
        """
        return self.execute(Command.GET_AVAILABLE_LOG_TYPES)["value"]

    def get_log(self, log_type):
        """Gets the log for a given log type.

        Parameters:
        -----------
        log_type : str
            - Type of log that which will be returned

        Example:
        --------
        >>> driver.get_log("browser")
        >>> driver.get_log("driver")
        >>> driver.get_log("client")
        >>> driver.get_log("server")
        """
        return self.execute(Command.GET_LOG, {"type": log_type})["value"]

    def set_sink_to_use(self, sink_name: str) -> dict:
        """Sets a specific sink, using its name, as a Cast session receiver
        target.

        :Args:
         - sink_name: Name of the sink to use as the target.
        """
        return self.execute("setSinkToUse", {"sinkName": sink_name})

    def start_desktop_mirroring(self, sink_name: str) -> dict:
        """Starts a desktop mirroring session on a specific receiver target.

        :Args:
         - sink_name: Name of the sink to use as the target.
        """
        return self.execute("startDesktopMirroring", {"sinkName": sink_name})

    def start_tab_mirroring(self, sink_name: str) -> dict:
        """Starts a tab mirroring session on a specific receiver target.

        :Args:
         - sink_name: Name of the sink to use as the target.
        """
        return self.execute("startTabMirroring", {"sinkName": sink_name})

    def stop_casting(self, sink_name: str) -> dict:
        """Stops the existing Cast session on a specific receiver target.

        :Args:
         - sink_name: Name of the sink to stop the Cast session.
        """
        return self.execute("stopCasting", {"sinkName": sink_name})

    def quit(self) -> None:
        """Closes the browser and shuts down the ChromiumDriver executable."""
        try:
            super().quit()
        except Exception:
            # We don't care about the message because something probably has gone wrong
            pass
        finally:
            self.service.stop()

    def download_file(self, *args, **kwargs):
        raise NotImplementedError

    def get_downloadable_files(self, *args, **kwargs):
        raise NotImplementedError

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\mechanics\models.py ===
#!/usr/bin/env python
"""This module contains some sample symbolic models used for testing and
examples."""

# Internal imports
from sympy.core import backend as sm
import sympy.physics.mechanics as me


def multi_mass_spring_damper(n=1, apply_gravity=False,
                             apply_external_forces=False):
    r"""Returns a system containing the symbolic equations of motion and
    associated variables for a simple multi-degree of freedom point mass,
    spring, damper system with optional gravitational and external
    specified forces. For example, a two mass system under the influence of
    gravity and external forces looks like:

    ::

        ----------------
         |     |     |   | g
         \    | |    |   V
      k0 /    --- c0 |
         |     |     | x0, v0
        ---------    V
        |  m0   | -----
        ---------    |
         | |   |     |
         \ v  | |    |
      k1 / f0 --- c1 |
         |     |     | x1, v1
        ---------    V
        |  m1   | -----
        ---------
           | f1
           V

    Parameters
    ==========

    n : integer
        The number of masses in the serial chain.
    apply_gravity : boolean
        If true, gravity will be applied to each mass.
    apply_external_forces : boolean
        If true, a time varying external force will be applied to each mass.

    Returns
    =======

    kane : sympy.physics.mechanics.kane.KanesMethod
        A KanesMethod object.

    """

    mass = sm.symbols('m:{}'.format(n))
    stiffness = sm.symbols('k:{}'.format(n))
    damping = sm.symbols('c:{}'.format(n))

    acceleration_due_to_gravity = sm.symbols('g')

    coordinates = me.dynamicsymbols('x:{}'.format(n))
    speeds = me.dynamicsymbols('v:{}'.format(n))
    specifieds = me.dynamicsymbols('f:{}'.format(n))

    ceiling = me.ReferenceFrame('N')
    origin = me.Point('origin')
    origin.set_vel(ceiling, 0)

    points = [origin]
    kinematic_equations = []
    particles = []
    forces = []

    for i in range(n):

        center = points[-1].locatenew('center{}'.format(i),
                                      coordinates[i] * ceiling.x)
        center.set_vel(ceiling, points[-1].vel(ceiling) +
                       speeds[i] * ceiling.x)
        points.append(center)

        block = me.Particle('block{}'.format(i), center, mass[i])

        kinematic_equations.append(speeds[i] - coordinates[i].diff())

        total_force = (-stiffness[i] * coordinates[i] -
                       damping[i] * speeds[i])
        try:
            total_force += (stiffness[i + 1] * coordinates[i + 1] +
                            damping[i + 1] * speeds[i + 1])
        except IndexError:  # no force from below on last mass
            pass

        if apply_gravity:
            total_force += mass[i] * acceleration_due_to_gravity

        if apply_external_forces:
            total_force += specifieds[i]

        forces.append((center, total_force * ceiling.x))

        particles.append(block)

    kane = me.KanesMethod(ceiling, q_ind=coordinates, u_ind=speeds,
                          kd_eqs=kinematic_equations)
    kane.kanes_equations(particles, forces)

    return kane


def n_link_pendulum_on_cart(n=1, cart_force=True, joint_torques=False):
    r"""Returns the system containing the symbolic first order equations of
    motion for a 2D n-link pendulum on a sliding cart under the influence of
    gravity.

    ::

                  |
         o    y   v
          \ 0 ^   g
           \  |
          --\-|----
          |  \|   |
      F-> |   o --|---> x
          |       |
          ---------
           o     o

    Parameters
    ==========

    n : integer
        The number of links in the pendulum.
    cart_force : boolean, default=True
        If true an external specified lateral force is applied to the cart.
    joint_torques : boolean, default=False
        If true joint torques will be added as specified inputs at each
        joint.

    Returns
    =======

    kane : sympy.physics.mechanics.kane.KanesMethod
        A KanesMethod object.

    Notes
    =====

    The degrees of freedom of the system are n + 1, i.e. one for each
    pendulum link and one for the lateral motion of the cart.

    M x' = F, where x = [u0, ..., un+1, q0, ..., qn+1]

    The joint angles are all defined relative to the ground where the x axis
    defines the ground line and the y axis points up. The joint torques are
    applied between each adjacent link and the between the cart and the
    lower link where a positive torque corresponds to positive angle.

    """
    if n <= 0:
        raise ValueError('The number of links must be a positive integer.')

    q = me.dynamicsymbols('q:{}'.format(n + 1))
    u = me.dynamicsymbols('u:{}'.format(n + 1))

    if joint_torques is True:
        T = me.dynamicsymbols('T1:{}'.format(n + 1))

    m = sm.symbols('m:{}'.format(n + 1))
    l = sm.symbols('l:{}'.format(n))
    g, t = sm.symbols('g t')

    I = me.ReferenceFrame('I')
    O = me.Point('O')
    O.set_vel(I, 0)

    P0 = me.Point('P0')
    P0.set_pos(O, q[0] * I.x)
    P0.set_vel(I, u[0] * I.x)
    Pa0 = me.Particle('Pa0', P0, m[0])

    frames = [I]
    points = [P0]
    particles = [Pa0]
    forces = [(P0, -m[0] * g * I.y)]
    kindiffs = [q[0].diff(t) - u[0]]

    if cart_force is True or joint_torques is True:
        specified = []
    else:
        specified = None

    for i in range(n):
        Bi = I.orientnew('B{}'.format(i), 'Axis', [q[i + 1], I.z])
        Bi.set_ang_vel(I, u[i + 1] * I.z)
        frames.append(Bi)

        Pi = points[-1].locatenew('P{}'.format(i + 1), l[i] * Bi.y)
        Pi.v2pt_theory(points[-1], I, Bi)
        points.append(Pi)

        Pai = me.Particle('Pa' + str(i + 1), Pi, m[i + 1])
        particles.append(Pai)

        forces.append((Pi, -m[i + 1] * g * I.y))

        if joint_torques is True:

            specified.append(T[i])

            if i == 0:
                forces.append((I, -T[i] * I.z))

            if i == n - 1:
                forces.append((Bi, T[i] * I.z))
            else:
                forces.append((Bi, T[i] * I.z - T[i + 1] * I.z))

        kindiffs.append(q[i + 1].diff(t) - u[i + 1])

    if cart_force is True:
        F = me.dynamicsymbols('F')
        forces.append((P0, F * I.x))
        specified.append(F)

    kane = me.KanesMethod(I, q_ind=q, u_ind=u, kd_eqs=kindiffs)
    kane.kanes_equations(particles, forces)

    return kane

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\quantum\trace.py ===
from sympy.core.add import Add
from sympy.core.containers import Tuple
from sympy.core.expr import Expr
from sympy.core.mul import Mul
from sympy.core.power import Pow
from sympy.core.sorting import default_sort_key
from sympy.core.sympify import sympify
from sympy.matrices import Matrix


def _is_scalar(e):
    """ Helper method used in Tr"""

    # sympify to set proper attributes
    e = sympify(e)
    if isinstance(e, Expr):
        if (e.is_Integer or e.is_Float or
            e.is_Rational or e.is_Number or
            (e.is_Symbol and e.is_commutative)
                ):
            return True

    return False


def _cycle_permute(l):
    """ Cyclic permutations based on canonical ordering

    Explanation
    ===========

    This method does the sort based ascii values while
    a better approach would be to used lexicographic sort.

    TODO: Handle condition such as symbols have subscripts/superscripts
    in case of lexicographic sort

    """

    if len(l) == 1:
        return l

    min_item = min(l, key=default_sort_key)
    indices = [i for i, x in enumerate(l) if x == min_item]

    le = list(l)
    le.extend(l)  # duplicate and extend string for easy processing

    # adding the first min_item index back for easier looping
    indices.append(len(l) + indices[0])

    # create sublist of items with first item as min_item and last_item
    # in each of the sublist is item just before the next occurrence of
    # minitem in the cycle formed.
    sublist = [[le[indices[i]:indices[i + 1]]] for i in
               range(len(indices) - 1)]

    # we do comparison of strings by comparing elements
    # in each sublist
    idx = sublist.index(min(sublist))
    ordered_l = le[indices[idx]:indices[idx] + len(l)]

    return ordered_l


def _rearrange_args(l):
    """ this just moves the last arg to first position
     to enable expansion of args
     A,B,A ==> A**2,B
    """
    if len(l) == 1:
        return l

    x = list(l[-1:])
    x.extend(l[0:-1])
    return Mul(*x).args


class Tr(Expr):
    """ Generic Trace operation than can trace over:

    a) SymPy matrix
    b) operators
    c) outer products

    Parameters
    ==========
    o : operator, matrix, expr
    i : tuple/list indices (optional)

    Examples
    ========

    # TODO: Need to handle printing

    a) Trace(A+B) = Tr(A) + Tr(B)
    b) Trace(scalar*Operator) = scalar*Trace(Operator)

    >>> from sympy.physics.quantum.trace import Tr
    >>> from sympy import symbols, Matrix
    >>> a, b = symbols('a b', commutative=True)
    >>> A, B = symbols('A B', commutative=False)
    >>> Tr(a*A,[2])
    a*Tr(A)
    >>> m = Matrix([[1,2],[1,1]])
    >>> Tr(m)
    2

    """
    def __new__(cls, *args):
        """ Construct a Trace object.

        Parameters
        ==========
        args = SymPy expression
        indices = tuple/list if indices, optional

        """

        # expect no indices,int or a tuple/list/Tuple
        if (len(args) == 2):
            if not isinstance(args[1], (list, Tuple, tuple)):
                indices = Tuple(args[1])
            else:
                indices = Tuple(*args[1])

            expr = args[0]
        elif (len(args) == 1):
            indices = Tuple()
            expr = args[0]
        else:
            raise ValueError("Arguments to Tr should be of form "
                             "(expr[, [indices]])")

        if isinstance(expr, Matrix):
            return expr.trace()
        elif hasattr(expr, 'trace') and callable(expr.trace):
            #for any objects that have trace() defined e.g numpy
            return expr.trace()
        elif isinstance(expr, Add):
            return Add(*[Tr(arg, indices) for arg in expr.args])
        elif isinstance(expr, Mul):
            c_part, nc_part = expr.args_cnc()
            if len(nc_part) == 0:
                return Mul(*c_part)
            else:
                obj = Expr.__new__(cls, Mul(*nc_part), indices )
                #this check is needed to prevent cached instances
                #being returned even if len(c_part)==0
                return Mul(*c_part)*obj if len(c_part) > 0 else obj
        elif isinstance(expr, Pow):
            if (_is_scalar(expr.args[0]) and
                    _is_scalar(expr.args[1])):
                return expr
            else:
                return Expr.__new__(cls, expr, indices)
        else:
            if (_is_scalar(expr)):
                return expr

            return Expr.__new__(cls, expr, indices)

    @property
    def kind(self):
        expr = self.args[0]
        expr_kind = expr.kind
        return expr_kind.element_kind

    def doit(self, **hints):
        """ Perform the trace operation.

        #TODO: Current version ignores the indices set for partial trace.

        >>> from sympy.physics.quantum.trace import Tr
        >>> from sympy.physics.quantum.operator import OuterProduct
        >>> from sympy.physics.quantum.spin import JzKet, JzBra
        >>> t = Tr(OuterProduct(JzKet(1,1), JzBra(1,1)))
        >>> t.doit()
        1

        """
        if hasattr(self.args[0], '_eval_trace'):
            return self.args[0]._eval_trace(indices=self.args[1])

        return self

    @property
    def is_number(self):
        # TODO : improve this implementation
        return True

    #TODO: Review if the permute method is needed
    # and if it needs to return a new instance
    def permute(self, pos):
        """ Permute the arguments cyclically.

        Parameters
        ==========

        pos : integer, if positive, shift-right, else shift-left

        Examples
        ========

        >>> from sympy.physics.quantum.trace import Tr
        >>> from sympy import symbols
        >>> A, B, C, D = symbols('A B C D', commutative=False)
        >>> t = Tr(A*B*C*D)
        >>> t.permute(2)
        Tr(C*D*A*B)
        >>> t.permute(-2)
        Tr(C*D*A*B)

        """
        if pos > 0:
            pos = pos % len(self.args[0].args)
        else:
            pos = -(abs(pos) % len(self.args[0].args))

        args = list(self.args[0].args[-pos:] + self.args[0].args[0:-pos])

        return Tr(Mul(*(args)))

    def _hashable_content(self):
        if isinstance(self.args[0], Mul):
            args = _cycle_permute(_rearrange_args(self.args[0].args))
        else:
            args = [self.args[0]]

        return tuple(args) + (self.args[1], )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\polys\matrices\linsolve.py ===
#
# sympy.polys.matrices.linsolve module
#
# This module defines the _linsolve function which is the internal workhorse
# used by linsolve. This computes the solution of a system of linear equations
# using the SDM sparse matrix implementation in sympy.polys.matrices.sdm. This
# is a replacement for solve_lin_sys in sympy.polys.solvers which is
# inefficient for large sparse systems due to the use of a PolyRing with many
# generators:
#
#     https://github.com/sympy/sympy/issues/20857
#
# The implementation of _linsolve here handles:
#
# - Extracting the coefficients from the Expr/Eq input equations.
# - Constructing a domain and converting the coefficients to
#   that domain.
# - Using the SDM.rref, SDM.nullspace etc methods to generate the full
#   solution working with arithmetic only in the domain of the coefficients.
#
# The routines here are particularly designed to be efficient for large sparse
# systems of linear equations although as well as dense systems. It is
# possible that for some small dense systems solve_lin_sys which uses the
# dense matrix implementation DDM will be more efficient. With smaller systems
# though the bulk of the time is spent just preprocessing the inputs and the
# relative time spent in rref is too small to be noticeable.
#

from collections import defaultdict

from sympy.core.add import Add
from sympy.core.mul import Mul
from sympy.core.singleton import S

from sympy.polys.constructor import construct_domain
from sympy.polys.solvers import PolyNonlinearError

from .sdm import (
    SDM,
    sdm_irref,
    sdm_particular_from_rref,
    sdm_nullspace_from_rref
)

from sympy.utilities.misc import filldedent


def _linsolve(eqs, syms):

    """Solve a linear system of equations.

    Examples
    ========

    Solve a linear system with a unique solution:

    >>> from sympy import symbols, Eq
    >>> from sympy.polys.matrices.linsolve import _linsolve
    >>> x, y = symbols('x, y')
    >>> eqs = [Eq(x + y, 1), Eq(x - y, 2)]
    >>> _linsolve(eqs, [x, y])
    {x: 3/2, y: -1/2}

    In the case of underdetermined systems the solution will be expressed in
    terms of the unknown symbols that are unconstrained:

    >>> _linsolve([Eq(x + y, 0)], [x, y])
    {x: -y, y: y}

    """
    # Number of unknowns (columns in the non-augmented matrix)
    nsyms = len(syms)

    # Convert to sparse augmented matrix (len(eqs) x (nsyms+1))
    eqsdict, const = _linear_eq_to_dict(eqs, syms)
    Aaug = sympy_dict_to_dm(eqsdict, const, syms)
    K = Aaug.domain

    # sdm_irref has issues with float matrices. This uses the ddm_rref()
    # function. When sdm_rref() can handle float matrices reasonably this
    # should be removed...
    if K.is_RealField or K.is_ComplexField:
        Aaug = Aaug.to_ddm().rref()[0].to_sdm()

    # Compute reduced-row echelon form (RREF)
    Arref, pivots, nzcols = sdm_irref(Aaug)

    # No solution:
    if pivots and pivots[-1] == nsyms:
        return None

    # Particular solution for non-homogeneous system:
    P = sdm_particular_from_rref(Arref, nsyms+1, pivots)

    # Nullspace - general solution to homogeneous system
    # Note: using nsyms not nsyms+1 to ignore last column
    V, nonpivots = sdm_nullspace_from_rref(Arref, K.one, nsyms, pivots, nzcols)

    # Collect together terms from particular and nullspace:
    sol = defaultdict(list)
    for i, v in P.items():
        sol[syms[i]].append(K.to_sympy(v))
    for npi, Vi in zip(nonpivots, V):
        sym = syms[npi]
        for i, v in Vi.items():
            sol[syms[i]].append(sym * K.to_sympy(v))

    # Use a single call to Add for each term:
    sol = {s: Add(*terms) for s, terms in sol.items()}

    # Fill in the zeros:
    zero = S.Zero
    for s in set(syms) - set(sol):
        sol[s] = zero

    # All done!
    return sol


def sympy_dict_to_dm(eqs_coeffs, eqs_rhs, syms):
    """Convert a system of dict equations to a sparse augmented matrix"""
    elems = set(eqs_rhs).union(*(e.values() for e in eqs_coeffs))
    K, elems_K = construct_domain(elems, field=True, extension=True)
    elem_map = dict(zip(elems, elems_K))
    neqs = len(eqs_coeffs)
    nsyms = len(syms)
    sym2index = dict(zip(syms, range(nsyms)))
    eqsdict = []
    for eq, rhs in zip(eqs_coeffs, eqs_rhs):
        eqdict = {sym2index[s]: elem_map[c] for s, c in eq.items()}
        if rhs:
            eqdict[nsyms] = -elem_map[rhs]
        if eqdict:
            eqsdict.append(eqdict)
    sdm_aug = SDM(enumerate(eqsdict), (neqs, nsyms + 1), K)
    return sdm_aug


def _linear_eq_to_dict(eqs, syms):
    """Convert a system Expr/Eq equations into dict form, returning
    the coefficient dictionaries and a list of syms-independent terms
    from each expression in ``eqs```.

    Examples
    ========

    >>> from sympy.polys.matrices.linsolve import _linear_eq_to_dict
    >>> from sympy.abc import x
    >>> _linear_eq_to_dict([2*x + 3], {x})
    ([{x: 2}], [3])
    """
    coeffs = []
    ind = []
    symset = set(syms)
    for e in eqs:
        if e.is_Equality:
            coeff, terms = _lin_eq2dict(e.lhs, symset)
            cR, tR = _lin_eq2dict(e.rhs, symset)
            # there were no nonlinear errors so now
            # cancellation is allowed
            coeff -= cR
            for k, v in tR.items():
                if k in terms:
                    terms[k] -= v
                else:
                    terms[k] = -v
            # don't store coefficients of 0, however
            terms = {k: v for k, v in terms.items() if v}
            c, d = coeff, terms
        else:
            c, d = _lin_eq2dict(e, symset)
        coeffs.append(d)
        ind.append(c)
    return coeffs, ind


def _lin_eq2dict(a, symset):
    """return (c, d) where c is the sym-independent part of ``a`` and
    ``d`` is an efficiently calculated dictionary mapping symbols to
    their coefficients. A PolyNonlinearError is raised if non-linearity
    is detected.

    The values in the dictionary will be non-zero.

    Examples
    ========

    >>> from sympy.polys.matrices.linsolve import _lin_eq2dict
    >>> from sympy.abc import x, y
    >>> _lin_eq2dict(x + 2*y + 3, {x, y})
    (3, {x: 1, y: 2})
    """
    if a in symset:
        return S.Zero, {a: S.One}
    elif a.is_Add:
        terms_list = defaultdict(list)
        coeff_list = []
        for ai in a.args:
            ci, ti = _lin_eq2dict(ai, symset)
            coeff_list.append(ci)
            for mij, cij in ti.items():
                terms_list[mij].append(cij)
        coeff = Add(*coeff_list)
        terms = {sym: Add(*coeffs) for sym, coeffs in terms_list.items()}
        return coeff, terms
    elif a.is_Mul:
        terms = terms_coeff = None
        coeff_list = []
        for ai in a.args:
            ci, ti = _lin_eq2dict(ai, symset)
            if not ti:
                coeff_list.append(ci)
            elif terms is None:
                terms = ti
                terms_coeff = ci
            else:
                # since ti is not null and we already have
                # a term, this is a cross term
                raise PolyNonlinearError(filldedent('''
                    nonlinear cross-term: %s''' % a))
        coeff = Mul._from_args(coeff_list)
        if terms is None:
            return coeff, {}
        else:
            terms = {sym: coeff * c for sym, c in terms.items()}
            return  coeff * terms_coeff, terms
    elif not a.has_xfree(symset):
        return a, {}
    else:
        raise PolyNonlinearError('nonlinear term: %s' % a)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\tomli_w\_writer.py ===
from __future__ import annotations

from collections.abc import Mapping
from datetime import date, datetime, time
from types import MappingProxyType

TYPE_CHECKING = False
if TYPE_CHECKING:
    from collections.abc import Generator
    from decimal import Decimal
    from typing import IO, Any, Final

ASCII_CTRL = frozenset(chr(i) for i in range(32)) | frozenset(chr(127))
ILLEGAL_BASIC_STR_CHARS = frozenset('"\\') | ASCII_CTRL - frozenset("\t")
BARE_KEY_CHARS = frozenset(
    "abcdefghijklmnopqrstuvwxyz" "ABCDEFGHIJKLMNOPQRSTUVWXYZ" "0123456789" "-_"
)
ARRAY_TYPES = (list, tuple)
MAX_LINE_LENGTH = 100

COMPACT_ESCAPES = MappingProxyType(
    {
        "\u0008": "\\b",  # backspace
        "\u000A": "\\n",  # linefeed
        "\u000C": "\\f",  # form feed
        "\u000D": "\\r",  # carriage return
        "\u0022": '\\"',  # quote
        "\u005C": "\\\\",  # backslash
    }
)


class Context:
    def __init__(self, allow_multiline: bool, indent: int):
        if indent < 0:
            raise ValueError("Indent width must be non-negative")
        self.allow_multiline: Final = allow_multiline
        # cache rendered inline tables (mapping from object id to rendered inline table)
        self.inline_table_cache: Final[dict[int, str]] = {}
        self.indent_str: Final = " " * indent


def dump(
    obj: Mapping[str, Any],
    fp: IO[bytes],
    /,
    *,
    multiline_strings: bool = False,
    indent: int = 4,
) -> None:
    ctx = Context(multiline_strings, indent)
    for chunk in gen_table_chunks(obj, ctx, name=""):
        fp.write(chunk.encode())


def dumps(
    obj: Mapping[str, Any], /, *, multiline_strings: bool = False, indent: int = 4
) -> str:
    ctx = Context(multiline_strings, indent)
    return "".join(gen_table_chunks(obj, ctx, name=""))


def gen_table_chunks(
    table: Mapping[str, Any],
    ctx: Context,
    *,
    name: str,
    inside_aot: bool = False,
) -> Generator[str, None, None]:
    yielded = False
    literals = []
    tables: list[tuple[str, Any, bool]] = []  # => [(key, value, inside_aot)]
    for k, v in table.items():
        if isinstance(v, Mapping):
            tables.append((k, v, False))
        elif is_aot(v) and not all(is_suitable_inline_table(t, ctx) for t in v):
            tables.extend((k, t, True) for t in v)
        else:
            literals.append((k, v))

    if inside_aot or name and (literals or not tables):
        yielded = True
        yield f"[[{name}]]\n" if inside_aot else f"[{name}]\n"

    if literals:
        yielded = True
        for k, v in literals:
            yield f"{format_key_part(k)} = {format_literal(v, ctx)}\n"

    for k, v, in_aot in tables:
        if yielded:
            yield "\n"
        else:
            yielded = True
        key_part = format_key_part(k)
        display_name = f"{name}.{key_part}" if name else key_part
        yield from gen_table_chunks(v, ctx, name=display_name, inside_aot=in_aot)


def format_literal(obj: object, ctx: Context, *, nest_level: int = 0) -> str:
    if isinstance(obj, bool):
        return "true" if obj else "false"
    if isinstance(obj, (int, float, date, datetime)):
        return str(obj)
    if isinstance(obj, time):
        if obj.tzinfo:
            raise ValueError("TOML does not support offset times")
        return str(obj)
    if isinstance(obj, str):
        return format_string(obj, allow_multiline=ctx.allow_multiline)
    if isinstance(obj, ARRAY_TYPES):
        return format_inline_array(obj, ctx, nest_level)
    if isinstance(obj, Mapping):
        return format_inline_table(obj, ctx)

    # Lazy import to improve module import time
    from decimal import Decimal

    if isinstance(obj, Decimal):
        return format_decimal(obj)
    raise TypeError(
        f"Object of type '{type(obj).__qualname__}' is not TOML serializable"
    )


def format_decimal(obj: Decimal) -> str:
    if obj.is_nan():
        return "nan"
    if obj.is_infinite():
        return "-inf" if obj.is_signed() else "inf"
    dec_str = str(obj).lower()
    return dec_str if "." in dec_str or "e" in dec_str else dec_str + ".0"


def format_inline_table(obj: Mapping, ctx: Context) -> str:
    # check cache first
    obj_id = id(obj)
    if obj_id in ctx.inline_table_cache:
        return ctx.inline_table_cache[obj_id]

    if not obj:
        rendered = "{}"
    else:
        rendered = (
            "{ "
            + ", ".join(
                f"{format_key_part(k)} = {format_literal(v, ctx)}"
                for k, v in obj.items()
            )
            + " }"
        )
    ctx.inline_table_cache[obj_id] = rendered
    return rendered


def format_inline_array(obj: tuple | list, ctx: Context, nest_level: int) -> str:
    if not obj:
        return "[]"
    item_indent = ctx.indent_str * (1 + nest_level)
    closing_bracket_indent = ctx.indent_str * nest_level
    return (
        "[\n"
        + ",\n".join(
            item_indent + format_literal(item, ctx, nest_level=nest_level + 1)
            for item in obj
        )
        + f",\n{closing_bracket_indent}]"
    )


def format_key_part(part: str) -> str:
    try:
        only_bare_key_chars = BARE_KEY_CHARS.issuperset(part)
    except TypeError:
        raise TypeError(
            f"Invalid mapping key '{part}' of type '{type(part).__qualname__}'."
            " A string is required."
        ) from None

    if part and only_bare_key_chars:
        return part
    return format_string(part, allow_multiline=False)


def format_string(s: str, *, allow_multiline: bool) -> str:
    do_multiline = allow_multiline and "\n" in s
    if do_multiline:
        result = '"""\n'
        s = s.replace("\r\n", "\n")
    else:
        result = '"'

    pos = seq_start = 0
    while True:
        try:
            char = s[pos]
        except IndexError:
            result += s[seq_start:pos]
            if do_multiline:
                return result + '"""'
            return result + '"'
        if char in ILLEGAL_BASIC_STR_CHARS:
            result += s[seq_start:pos]
            if char in COMPACT_ESCAPES:
                if do_multiline and char == "\n":
                    result += "\n"
                else:
                    result += COMPACT_ESCAPES[char]
            else:
                result += "\\u" + hex(ord(char))[2:].rjust(4, "0")
            seq_start = pos + 1
        pos += 1


def is_aot(obj: Any) -> bool:
    """Decides if an object behaves as an array of tables (i.e. a nonempty list
    of dicts)."""
    return bool(
        isinstance(obj, ARRAY_TYPES)
        and obj
        and all(isinstance(v, Mapping) for v in obj)
    )


def is_suitable_inline_table(obj: Mapping, ctx: Context) -> bool:
    """Use heuristics to decide if the inline-style representation is a good
    choice for a given table."""
    rendered_inline = f"{ctx.indent_str}{format_inline_table(obj, ctx)},"
    return len(rendered_inline) <= MAX_LINE_LENGTH and "\n" not in rendered_inline

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\firefox\webdriver.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.
import base64
import os
import warnings
import zipfile
from contextlib import contextmanager
from io import BytesIO
from typing import Optional

from selenium.webdriver.common.driver_finder import DriverFinder
from selenium.webdriver.remote.webdriver import WebDriver as RemoteWebDriver

from .options import Options
from .remote_connection import FirefoxRemoteConnection
from .service import Service


class WebDriver(RemoteWebDriver):
    """Controls the GeckoDriver and allows you to drive the browser."""

    CONTEXT_CHROME = "chrome"
    CONTEXT_CONTENT = "content"

    def __init__(
        self,
        options: Optional[Options] = None,
        service: Optional[Service] = None,
        keep_alive: bool = True,
    ) -> None:
        """Creates a new instance of the Firefox driver. Starts the service and
        then creates new instance of Firefox driver.

        :Args:
         - options - Instance of ``options.Options``.
         - service - (Optional) service instance for managing the starting and stopping of the driver.
         - keep_alive - Whether to configure remote_connection.RemoteConnection to use HTTP keep-alive.
        """

        self.service = service if service else Service()
        options = options if options else Options()

        finder = DriverFinder(self.service, options)
        if finder.get_browser_path():
            options.binary_location = finder.get_browser_path()
            options.browser_version = None

        self.service.path = self.service.env_path() or finder.get_driver_path()
        self.service.start()

        executor = FirefoxRemoteConnection(
            remote_server_addr=self.service.service_url,
            keep_alive=keep_alive,
            ignore_proxy=options._ignore_local_proxy,
        )

        try:
            super().__init__(command_executor=executor, options=options)
        except Exception:
            self.quit()
            raise

        self._is_remote = False

    def quit(self) -> None:
        """Closes the browser and shuts down the GeckoDriver executable."""
        try:
            super().quit()
        except Exception:
            # We don't care about the message because something probably has gone wrong
            pass
        finally:
            self.service.stop()

    def set_context(self, context) -> None:
        self.execute("SET_CONTEXT", {"context": context})

    @contextmanager
    def context(self, context):
        """Sets the context that Selenium commands are running in using a
        `with` statement. The state of the context on the server is saved
        before entering the block, and restored upon exiting it.

        :param context: Context, may be one of the class properties
            `CONTEXT_CHROME` or `CONTEXT_CONTENT`.

        Usage example::

            with selenium.context(selenium.CONTEXT_CHROME):
                # chrome scope
                ... do stuff ...
        """
        initial_context = self.execute("GET_CONTEXT").pop("value")
        self.set_context(context)
        try:
            yield
        finally:
            self.set_context(initial_context)

    def install_addon(self, path, temporary=False) -> str:
        """Installs Firefox addon.

        Returns identifier of installed addon. This identifier can later
        be used to uninstall addon.

        :param temporary: allows you to load browser extensions temporarily during a session
        :param path: Absolute path to the addon that will be installed.

        :Usage:
            ::

                driver.install_addon("/path/to/firebug.xpi")
        """

        if os.path.isdir(path):
            fp = BytesIO()
            # filter all trailing slash found in path
            path = os.path.normpath(path)
            # account for trailing slash that will be added by os.walk()
            path_root = len(path) + 1
            with zipfile.ZipFile(fp, "w", zipfile.ZIP_DEFLATED, strict_timestamps=False) as zipped:
                for base, _, files in os.walk(path):
                    for fyle in files:
                        filename = os.path.join(base, fyle)
                        zipped.write(filename, filename[path_root:])
            addon = base64.b64encode(fp.getvalue()).decode("UTF-8")
        else:
            with open(path, "rb") as file:
                addon = base64.b64encode(file.read()).decode("UTF-8")

        payload = {"addon": addon, "temporary": temporary}
        return self.execute("INSTALL_ADDON", payload)["value"]

    def uninstall_addon(self, identifier) -> None:
        """Uninstalls Firefox addon using its identifier.

        :Usage:
            ::

                driver.uninstall_addon("addon@foo.com")
        """
        self.execute("UNINSTALL_ADDON", {"id": identifier})

    def get_full_page_screenshot_as_file(self, filename) -> bool:
        """Saves a full document screenshot of the current window to a PNG
        image file. Returns False if there is any IOError, else returns True.
        Use full paths in your filename.

        :Args:
         - filename: The full path you wish to save your screenshot to. This
           should end with a `.png` extension.

        :Usage:
            ::

                driver.get_full_page_screenshot_as_file("/Screenshots/foo.png")
        """
        if not filename.lower().endswith(".png"):
            warnings.warn(
                "name used for saved screenshot does not match file type. It should end with a `.png` extension",
                UserWarning,
            )
        png = self.get_full_page_screenshot_as_png()
        try:
            with open(filename, "wb") as f:
                f.write(png)
        except OSError:
            return False
        finally:
            del png
        return True

    def save_full_page_screenshot(self, filename) -> bool:
        """Saves a full document screenshot of the current window to a PNG
        image file. Returns False if there is any IOError, else returns True.
        Use full paths in your filename.

        :Args:
         - filename: The full path you wish to save your screenshot to. This
           should end with a `.png` extension.

        :Usage:
            ::

                driver.save_full_page_screenshot("/Screenshots/foo.png")
        """
        return self.get_full_page_screenshot_as_file(filename)

    def get_full_page_screenshot_as_png(self) -> bytes:
        """Gets the full document screenshot of the current window as a binary
        data.

        :Usage:
            ::

                driver.get_full_page_screenshot_as_png()
        """
        return base64.b64decode(self.get_full_page_screenshot_as_base64().encode("ascii"))

    def get_full_page_screenshot_as_base64(self) -> str:
        """Gets the full document screenshot of the current window as a base64
        encoded string which is useful in embedded images in HTML.

        :Usage:
            ::

                driver.get_full_page_screenshot_as_base64()
        """
        return self.execute("FULL_PAGE_SCREENSHOT")["value"]

    def download_file(self, *args, **kwargs):
        raise NotImplementedError

    def get_downloadable_files(self, *args, **kwargs):
        raise NotImplementedError

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\type_tests\raisesgroup.py ===
from __future__ import annotations

import sys
from typing import Callable, Union

from trio.testing import Matcher, RaisesGroup
from typing_extensions import assert_type

if sys.version_info < (3, 11):
    from exceptiongroup import BaseExceptionGroup, ExceptionGroup

# split into functions to isolate the different scopes


def check_matcher_typevar_default(e: Matcher) -> None:
    assert e.exception_type is not None
    _exc: type[BaseException] = e.exception_type
    # this would previously pass, as the type would be `Any`
    e.exception_type().blah()  # type: ignore


def check_basic_contextmanager() -> None:
    with RaisesGroup(ValueError) as e:
        raise ExceptionGroup("foo", (ValueError(),))
    assert_type(e.value, ExceptionGroup[ValueError])


def check_basic_matches() -> None:
    # check that matches gets rid of the naked ValueError in the union
    exc: ExceptionGroup[ValueError] | ValueError = ExceptionGroup("", (ValueError(),))
    if RaisesGroup(ValueError).matches(exc):
        assert_type(exc, ExceptionGroup[ValueError])

    # also check that BaseExceptionGroup shows up for BaseExceptions
    if RaisesGroup(KeyboardInterrupt).matches(exc):
        assert_type(exc, BaseExceptionGroup[KeyboardInterrupt])


def check_matches_with_different_exception_type() -> None:
    e: BaseExceptionGroup[KeyboardInterrupt] = BaseExceptionGroup(
        "",
        (KeyboardInterrupt(),),
    )

    # note: it might be tempting to have this warn.
    # however, that isn't possible with current typing
    if RaisesGroup(ValueError).matches(e):
        assert_type(e, ExceptionGroup[ValueError])


def check_matcher_init() -> None:
    def check_exc(exc: BaseException) -> bool:
        return isinstance(exc, ValueError)

    # Check various combinations of constructor signatures.
    # At least 1 arg must be provided.
    Matcher()  # type: ignore
    Matcher(ValueError)
    Matcher(ValueError, "regex")
    Matcher(ValueError, "regex", check_exc)
    Matcher(exception_type=ValueError)
    Matcher(match="regex")
    Matcher(check=check_exc)
    Matcher(ValueError, match="regex")
    Matcher(match="regex", check=check_exc)

    def check_filenotfound(exc: FileNotFoundError) -> bool:
        return not exc.filename.endswith(".tmp")

    # If exception_type is provided, that narrows the `check` method's argument.
    Matcher(FileNotFoundError, check=check_filenotfound)
    Matcher(ValueError, check=check_filenotfound)  # type: ignore
    Matcher(check=check_filenotfound)  # type: ignore
    Matcher(FileNotFoundError, match="regex", check=check_filenotfound)


def raisesgroup_check_type_narrowing() -> None:
    """Check type narrowing on the `check` argument to `RaisesGroup`.
    All `type: ignore`s are correctly pointing out type errors.
    """

    def handle_exc(e: BaseExceptionGroup[BaseException]) -> bool:
        return True

    def handle_kbi(e: BaseExceptionGroup[KeyboardInterrupt]) -> bool:
        return True

    def handle_value(e: BaseExceptionGroup[ValueError]) -> bool:
        return True

    RaisesGroup(BaseException, check=handle_exc)
    RaisesGroup(BaseException, check=handle_kbi)  # type: ignore

    RaisesGroup(Exception, check=handle_exc)
    RaisesGroup(Exception, check=handle_value)  # type: ignore

    RaisesGroup(KeyboardInterrupt, check=handle_exc)
    RaisesGroup(KeyboardInterrupt, check=handle_kbi)
    RaisesGroup(KeyboardInterrupt, check=handle_value)  # type: ignore

    RaisesGroup(ValueError, check=handle_exc)
    RaisesGroup(ValueError, check=handle_kbi)  # type: ignore
    RaisesGroup(ValueError, check=handle_value)

    RaisesGroup(ValueError, KeyboardInterrupt, check=handle_exc)
    RaisesGroup(ValueError, KeyboardInterrupt, check=handle_kbi)  # type: ignore
    RaisesGroup(ValueError, KeyboardInterrupt, check=handle_value)  # type: ignore


def raisesgroup_narrow_baseexceptiongroup() -> None:
    """Check type narrowing specifically for the container exceptiongroup."""

    def handle_group(e: ExceptionGroup[Exception]) -> bool:
        return True

    def handle_group_value(e: ExceptionGroup[ValueError]) -> bool:
        return True

    RaisesGroup(ValueError, check=handle_group_value)

    RaisesGroup(Exception, check=handle_group)


def check_matcher_transparent() -> None:
    with RaisesGroup(Matcher(ValueError)) as e:
        ...
    _: BaseExceptionGroup[ValueError] = e.value
    assert_type(e.value, ExceptionGroup[ValueError])


def check_nested_raisesgroups_contextmanager() -> None:
    with RaisesGroup(RaisesGroup(ValueError)) as excinfo:
        raise ExceptionGroup("foo", (ValueError(),))

    _: BaseExceptionGroup[BaseExceptionGroup[ValueError]] = excinfo.value

    assert_type(
        excinfo.value,
        ExceptionGroup[ExceptionGroup[ValueError]],
    )

    assert_type(
        excinfo.value.exceptions[0],
        # this union is because of how typeshed defines .exceptions
        Union[
            ExceptionGroup[ValueError],
            ExceptionGroup[ExceptionGroup[ValueError]],
        ],
    )


def check_nested_raisesgroups_matches() -> None:
    """Check nested RaisesGroups with .matches"""
    exc: ExceptionGroup[ExceptionGroup[ValueError]] = ExceptionGroup(
        "",
        (ExceptionGroup("", (ValueError(),)),),
    )

    if RaisesGroup(RaisesGroup(ValueError)).matches(exc):
        assert_type(exc, ExceptionGroup[ExceptionGroup[ValueError]])


def check_multiple_exceptions_1() -> None:
    a = RaisesGroup(ValueError, ValueError)
    b = RaisesGroup(Matcher(ValueError), Matcher(ValueError))
    c = RaisesGroup(ValueError, Matcher(ValueError))

    d: RaisesGroup[ValueError]
    d = a
    d = b
    d = c
    assert isinstance(d, RaisesGroup)


def check_multiple_exceptions_2() -> None:
    # This previously failed due to lack of covariance in the TypeVar
    a = RaisesGroup(Matcher(ValueError), Matcher(TypeError))
    b = RaisesGroup(Matcher(ValueError), TypeError)
    c = RaisesGroup(ValueError, TypeError)

    d: RaisesGroup[Exception]
    d = a
    d = b
    d = c
    assert isinstance(d, RaisesGroup)


def check_raisesgroup_overloads() -> None:
    # allow_unwrapped=True does not allow:
    # multiple exceptions
    RaisesGroup(ValueError, TypeError, allow_unwrapped=True)  # type: ignore
    # nested RaisesGroup
    RaisesGroup(RaisesGroup(ValueError), allow_unwrapped=True)  # type: ignore
    # specifying match
    RaisesGroup(ValueError, match="foo", allow_unwrapped=True)  # type: ignore
    # specifying check
    RaisesGroup(ValueError, check=bool, allow_unwrapped=True)  # type: ignore
    # allowed variants
    RaisesGroup(ValueError, allow_unwrapped=True)
    RaisesGroup(ValueError, allow_unwrapped=True, flatten_subgroups=True)
    RaisesGroup(Matcher(ValueError), allow_unwrapped=True)

    # flatten_subgroups=True does not allow nested RaisesGroup
    RaisesGroup(RaisesGroup(ValueError), flatten_subgroups=True)  # type: ignore
    # but rest is plenty fine
    RaisesGroup(ValueError, TypeError, flatten_subgroups=True)
    RaisesGroup(ValueError, match="foo", flatten_subgroups=True)
    RaisesGroup(ValueError, check=bool, flatten_subgroups=True)
    RaisesGroup(ValueError, flatten_subgroups=True)
    RaisesGroup(Matcher(ValueError), flatten_subgroups=True)

    # if they're both false we can of course specify nested raisesgroup
    RaisesGroup(RaisesGroup(ValueError))


def check_triple_nested_raisesgroup() -> None:
    with RaisesGroup(RaisesGroup(RaisesGroup(ValueError))) as e:
        assert_type(e.value, ExceptionGroup[ExceptionGroup[ExceptionGroup[ValueError]]])


def check_check_typing() -> None:
    # `BaseExceptiongroup` should perhaps be `ExceptionGroup`, but close enough
    assert_type(
        RaisesGroup(ValueError).check,
        Union[
            Callable[[BaseExceptionGroup[ValueError]], bool],
            None,
        ],
    )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\wtforms\fields\choices.py ===
import itertools

from wtforms import widgets
from wtforms.fields.core import Field
from wtforms.validators import ValidationError

__all__ = (
    "SelectField",
    "SelectMultipleField",
    "RadioField",
)


class SelectFieldBase(Field):
    option_widget = widgets.Option()

    """
    Base class for fields which can be iterated to produce options.

    This isn't a field, but an abstract base class for fields which want to
    provide this functionality.
    """

    def __init__(self, label=None, validators=None, option_widget=None, **kwargs):
        super().__init__(label, validators, **kwargs)

        if option_widget is not None:
            self.option_widget = option_widget

    def iter_choices(self):
        """
        Provides data for choice widget rendering. Must return a sequence or
        iterable of (value, label, selected, render_kw) tuples.
        """
        raise NotImplementedError()

    def has_groups(self):
        return False

    def iter_groups(self):
        raise NotImplementedError()

    def __iter__(self):
        opts = dict(
            widget=self.option_widget,
            validators=self.validators,
            name=self.name,
            render_kw=self.render_kw,
            _form=None,
            _meta=self.meta,
        )
        for i, choice in enumerate(self.iter_choices()):
            if len(choice) == 4:
                value, label, checked, render_kw = choice
            else:
                value, label, checked = choice
                render_kw = {}

            opt = self._Option(
                label=label, id="%s-%d" % (self.id, i), **opts, **render_kw
            )
            opt.process(None, value)
            opt.checked = checked
            yield opt

    class _Option(Field):
        checked = False

        def _value(self):
            return str(self.data)


class SelectField(SelectFieldBase):
    widget = widgets.Select()

    def __init__(
        self,
        label=None,
        validators=None,
        coerce=str,
        choices=None,
        validate_choice=True,
        **kwargs,
    ):
        super().__init__(label, validators, **kwargs)
        self.coerce = coerce
        if callable(choices):
            choices = choices()
        if choices is not None:
            self.choices = choices if isinstance(choices, dict) else list(choices)
        else:
            self.choices = None
        self.validate_choice = validate_choice

    def iter_choices(self):
        if not self.choices:
            choices = []
        elif isinstance(self.choices, dict):
            choices = list(itertools.chain.from_iterable(self.choices.values()))
        else:
            choices = self.choices

        return self._choices_generator(choices)

    def has_groups(self):
        return isinstance(self.choices, dict)

    def iter_groups(self):
        if isinstance(self.choices, dict):
            for label, choices in self.choices.items():
                yield (label, self._choices_generator(choices))

    def _choices_generator(self, choices):
        if not choices:
            _choices = []

        elif isinstance(choices[0], (list, tuple)):
            _choices = choices

        else:
            _choices = zip(choices, choices)

        for value, label, *other_args in _choices:
            selected = self.coerce(value) == self.data
            render_kw = other_args[0] if len(other_args) else {}
            yield (value, label, selected, render_kw)

    def process_data(self, value):
        try:
            # If value is None, don't coerce to a value
            self.data = self.coerce(value) if value is not None else None
        except (ValueError, TypeError):
            self.data = None

    def process_formdata(self, valuelist):
        if not valuelist:
            return

        try:
            self.data = self.coerce(valuelist[0])
        except ValueError as exc:
            raise ValueError(self.gettext("Invalid Choice: could not coerce.")) from exc

    def pre_validate(self, form):
        if not self.validate_choice:
            return

        if self.choices is None:
            raise TypeError(self.gettext("Choices cannot be None."))

        for _, _, match, *_ in self.iter_choices():
            if match:
                break
        else:
            raise ValidationError(self.gettext("Not a valid choice."))


class SelectMultipleField(SelectField):
    """
    No different from a normal select field, except this one can take (and
    validate) multiple choices.  You'll need to specify the HTML `size`
    attribute to the select field when rendering.
    """

    widget = widgets.Select(multiple=True)

    def _choices_generator(self, choices):
        if not choices:
            _choices = []

        elif isinstance(choices[0], (list, tuple)):
            _choices = choices

        else:
            _choices = zip(choices, choices)

        for value, label, *other_args in _choices:
            selected = self.data is not None and self.coerce(value) in self.data
            render_kw = other_args[0] if len(other_args) else {}
            yield (value, label, selected, render_kw)

    def process_data(self, value):
        try:
            self.data = list(self.coerce(v) for v in value)
        except (ValueError, TypeError):
            self.data = None

    def process_formdata(self, valuelist):
        try:
            self.data = list(self.coerce(x) for x in valuelist)
        except ValueError as exc:
            raise ValueError(
                self.gettext(
                    "Invalid choice(s): one or more data inputs could not be coerced."
                )
            ) from exc

    def pre_validate(self, form):
        if not self.validate_choice or not self.data:
            return

        if self.choices is None:
            raise TypeError(self.gettext("Choices cannot be None."))

        acceptable = [self.coerce(choice[0]) for choice in self.iter_choices()]
        if any(data not in acceptable for data in self.data):
            unacceptable = [
                str(data) for data in set(self.data) if data not in acceptable
            ]
            raise ValidationError(
                self.ngettext(
                    "'%(value)s' is not a valid choice for this field.",
                    "'%(value)s' are not valid choices for this field.",
                    len(unacceptable),
                )
                % dict(value="', '".join(unacceptable))
            )


class RadioField(SelectField):
    """
    Like a SelectField, except displays a list of radio buttons.

    Iterating the field will produce subfields (each containing a label as
    well) in order to allow custom rendering of the individual radio fields.
    """

    widget = widgets.ListWidget(prefix_label=False)
    option_widget = widgets.RadioInput()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\itsdangerous\timed.py ===
from __future__ import annotations

import collections.abc as cabc
import time
import typing as t
from datetime import datetime
from datetime import timezone

from .encoding import base64_decode
from .encoding import base64_encode
from .encoding import bytes_to_int
from .encoding import int_to_bytes
from .encoding import want_bytes
from .exc import BadSignature
from .exc import BadTimeSignature
from .exc import SignatureExpired
from .serializer import _TSerialized
from .serializer import Serializer
from .signer import Signer


class TimestampSigner(Signer):
    """Works like the regular :class:`.Signer` but also records the time
    of the signing and can be used to expire signatures. The
    :meth:`unsign` method can raise :exc:`.SignatureExpired` if the
    unsigning failed because the signature is expired.
    """

    def get_timestamp(self) -> int:
        """Returns the current timestamp. The function must return an
        integer.
        """
        return int(time.time())

    def timestamp_to_datetime(self, ts: int) -> datetime:
        """Convert the timestamp from :meth:`get_timestamp` into an
        aware :class`datetime.datetime` in UTC.

        .. versionchanged:: 2.0
            The timestamp is returned as a timezone-aware ``datetime``
            in UTC rather than a naive ``datetime`` assumed to be UTC.
        """
        return datetime.fromtimestamp(ts, tz=timezone.utc)

    def sign(self, value: str | bytes) -> bytes:
        """Signs the given string and also attaches time information."""
        value = want_bytes(value)
        timestamp = base64_encode(int_to_bytes(self.get_timestamp()))
        sep = want_bytes(self.sep)
        value = value + sep + timestamp
        return value + sep + self.get_signature(value)

    # Ignore overlapping signatures check, return_timestamp is the only
    # parameter that affects the return type.

    @t.overload
    def unsign(  # type: ignore[overload-overlap]
        self,
        signed_value: str | bytes,
        max_age: int | None = None,
        return_timestamp: t.Literal[False] = False,
    ) -> bytes: ...

    @t.overload
    def unsign(
        self,
        signed_value: str | bytes,
        max_age: int | None = None,
        return_timestamp: t.Literal[True] = True,
    ) -> tuple[bytes, datetime]: ...

    def unsign(
        self,
        signed_value: str | bytes,
        max_age: int | None = None,
        return_timestamp: bool = False,
    ) -> tuple[bytes, datetime] | bytes:
        """Works like the regular :meth:`.Signer.unsign` but can also
        validate the time. See the base docstring of the class for
        the general behavior. If ``return_timestamp`` is ``True`` the
        timestamp of the signature will be returned as an aware
        :class:`datetime.datetime` object in UTC.

        .. versionchanged:: 2.0
            The timestamp is returned as a timezone-aware ``datetime``
            in UTC rather than a naive ``datetime`` assumed to be UTC.
        """
        try:
            result = super().unsign(signed_value)
            sig_error = None
        except BadSignature as e:
            sig_error = e
            result = e.payload or b""

        sep = want_bytes(self.sep)

        # If there is no timestamp in the result there is something
        # seriously wrong. In case there was a signature error, we raise
        # that one directly, otherwise we have a weird situation in
        # which we shouldn't have come except someone uses a time-based
        # serializer on non-timestamp data, so catch that.
        if sep not in result:
            if sig_error:
                raise sig_error

            raise BadTimeSignature("timestamp missing", payload=result)

        value, ts_bytes = result.rsplit(sep, 1)
        ts_int: int | None = None
        ts_dt: datetime | None = None

        try:
            ts_int = bytes_to_int(base64_decode(ts_bytes))
        except Exception:
            pass

        # Signature is *not* okay. Raise a proper error now that we have
        # split the value and the timestamp.
        if sig_error is not None:
            if ts_int is not None:
                try:
                    ts_dt = self.timestamp_to_datetime(ts_int)
                except (ValueError, OSError, OverflowError) as exc:
                    # Windows raises OSError
                    # 32-bit raises OverflowError
                    raise BadTimeSignature(
                        "Malformed timestamp", payload=value
                    ) from exc

            raise BadTimeSignature(str(sig_error), payload=value, date_signed=ts_dt)

        # Signature was okay but the timestamp is actually not there or
        # malformed. Should not happen, but we handle it anyway.
        if ts_int is None:
            raise BadTimeSignature("Malformed timestamp", payload=value)

        # Check timestamp is not older than max_age
        if max_age is not None:
            age = self.get_timestamp() - ts_int

            if age > max_age:
                raise SignatureExpired(
                    f"Signature age {age} > {max_age} seconds",
                    payload=value,
                    date_signed=self.timestamp_to_datetime(ts_int),
                )

            if age < 0:
                raise SignatureExpired(
                    f"Signature age {age} < 0 seconds",
                    payload=value,
                    date_signed=self.timestamp_to_datetime(ts_int),
                )

        if return_timestamp:
            return value, self.timestamp_to_datetime(ts_int)

        return value

    def validate(self, signed_value: str | bytes, max_age: int | None = None) -> bool:
        """Only validates the given signed value. Returns ``True`` if
        the signature exists and is valid."""
        try:
            self.unsign(signed_value, max_age=max_age)
            return True
        except BadSignature:
            return False


class TimedSerializer(Serializer[_TSerialized]):
    """Uses :class:`TimestampSigner` instead of the default
    :class:`.Signer`.
    """

    default_signer: type[TimestampSigner] = TimestampSigner

    def iter_unsigners(
        self, salt: str | bytes | None = None
    ) -> cabc.Iterator[TimestampSigner]:
        return t.cast("cabc.Iterator[TimestampSigner]", super().iter_unsigners(salt))

    # TODO: Signature is incompatible because parameters were added
    #  before salt.

    def loads(  # type: ignore[override]
        self,
        s: str | bytes,
        max_age: int | None = None,
        return_timestamp: bool = False,
        salt: str | bytes | None = None,
    ) -> t.Any:
        """Reverse of :meth:`dumps`, raises :exc:`.BadSignature` if the
        signature validation fails. If a ``max_age`` is provided it will
        ensure the signature is not older than that time in seconds. In
        case the signature is outdated, :exc:`.SignatureExpired` is
        raised. All arguments are forwarded to the signer's
        :meth:`~TimestampSigner.unsign` method.
        """
        s = want_bytes(s)
        last_exception = None

        for signer in self.iter_unsigners(salt):
            try:
                base64d, timestamp = signer.unsign(
                    s, max_age=max_age, return_timestamp=True
                )
                payload = self.load_payload(base64d)

                if return_timestamp:
                    return payload, timestamp

                return payload
            except SignatureExpired:
                # The signature was unsigned successfully but was
                # expired. Do not try the next signer.
                raise
            except BadSignature as err:
                last_exception = err

        raise t.cast(BadSignature, last_exception)

    def loads_unsafe(  # type: ignore[override]
        self,
        s: str | bytes,
        max_age: int | None = None,
        salt: str | bytes | None = None,
    ) -> tuple[bool, t.Any]:
        return self._loads_unsafe_impl(s, salt, load_kwargs={"max_age": max_age})

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\machine_info.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------

# It is used to dump machine information for Notebooks

import argparse
import json
import logging
import platform
from os import environ

import cpuinfo
import psutil
from py3nvml.py3nvml import (
    NVMLError,
    nvmlDeviceGetCount,
    nvmlDeviceGetHandleByIndex,
    nvmlDeviceGetMemoryInfo,
    nvmlDeviceGetName,
    nvmlInit,
    nvmlShutdown,
    nvmlSystemGetDriverVersion,
)


class MachineInfo:
    """Class encapsulating Machine Info logic."""

    def __init__(self, silent=False, logger=None):
        self.silent = silent

        if logger is None:
            logging.basicConfig(
                format="%(asctime)s - %(name)s - %(levelname)s: %(message)s",
                level=logging.INFO,
            )
            self.logger = logging.getLogger(__name__)
        else:
            self.logger = logger

        self.machine_info = None
        try:
            self.machine_info = self.get_machine_info()
        except Exception:
            self.logger.exception("Exception in getting machine info.")
            self.machine_info = None

    def get_machine_info(self):
        """Get machine info in metric format"""
        gpu_info = self.get_gpu_info_by_nvml()
        cpu_info = cpuinfo.get_cpu_info()

        machine_info = {
            "gpu": gpu_info,
            "cpu": self.get_cpu_info(),
            "memory": self.get_memory_info(),
            "os": platform.platform(),
            "python": self._try_get(cpu_info, ["python_version"]),
            "packages": self.get_related_packages(),
            "onnxruntime": self.get_onnxruntime_info(),
            "pytorch": self.get_pytorch_info(),
            "tensorflow": self.get_tensorflow_info(),
        }
        return machine_info

    def get_memory_info(self) -> dict:
        """Get memory info"""
        mem = psutil.virtual_memory()
        return {"total": mem.total, "available": mem.available}

    def _try_get(self, cpu_info: dict, names: list) -> str:
        for name in names:
            if name in cpu_info:
                value = cpu_info[name]
                if isinstance(value, (list, tuple)):
                    return ",".join([str(i) for i in value])
                return value
        return ""

    def get_cpu_info(self) -> dict:
        """Get CPU info"""
        cpu_info = cpuinfo.get_cpu_info()

        return {
            "brand": self._try_get(cpu_info, ["brand", "brand_raw"]),
            "cores": psutil.cpu_count(logical=False),
            "logical_cores": psutil.cpu_count(logical=True),
            "hz": self._try_get(cpu_info, ["hz_actual"]),
            "l2_cache": self._try_get(cpu_info, ["l2_cache_size"]),
            "flags": self._try_get(cpu_info, ["flags"]),
            "processor": platform.uname().processor,
        }

    def get_gpu_info_by_nvml(self) -> dict:
        """Get GPU info using nvml"""
        gpu_info_list = []
        driver_version = None
        try:
            nvmlInit()
            driver_version = nvmlSystemGetDriverVersion()
            deviceCount = nvmlDeviceGetCount()  # noqa: N806
            for i in range(deviceCount):
                handle = nvmlDeviceGetHandleByIndex(i)
                info = nvmlDeviceGetMemoryInfo(handle)
                gpu_info = {}
                gpu_info["memory_total"] = info.total
                gpu_info["memory_available"] = info.free
                gpu_info["name"] = nvmlDeviceGetName(handle)
                gpu_info_list.append(gpu_info)
            nvmlShutdown()
        except NVMLError as error:
            if not self.silent:
                self.logger.error("Error fetching GPU information using nvml: %s", error)
            return None

        result = {"driver_version": driver_version, "devices": gpu_info_list}

        if "CUDA_VISIBLE_DEVICES" in environ:
            result["cuda_visible"] = environ["CUDA_VISIBLE_DEVICES"]
        return result

    def get_related_packages(self) -> list[str]:
        import pkg_resources

        installed_packages = pkg_resources.working_set
        related_packages = [
            "onnxruntime-gpu",
            "onnxruntime",
            "onnx",
            "transformers",
            "protobuf",
            "sympy",
            "torch",
            "tensorflow",
            "flatbuffers",
            "numpy",
            "onnxconverter-common",
        ]
        related_packages_list = {i.key: i.version for i in installed_packages if i.key in related_packages}
        return related_packages_list

    def get_onnxruntime_info(self) -> dict:
        try:
            import onnxruntime

            return {
                "version": onnxruntime.__version__,
                "support_gpu": "CUDAExecutionProvider" in onnxruntime.get_available_providers(),
            }
        except ImportError as error:
            if not self.silent:
                self.logger.exception(error)
            return None
        except Exception as exception:
            if not self.silent:
                self.logger.exception(exception, False)
            return None

    def get_pytorch_info(self) -> dict:
        try:
            import torch

            return {
                "version": torch.__version__,
                "support_gpu": torch.cuda.is_available(),
                "cuda": torch.version.cuda,
            }
        except ImportError as error:
            if not self.silent:
                self.logger.exception(error)
            return None
        except Exception as exception:
            if not self.silent:
                self.logger.exception(exception, False)
            return None

    def get_tensorflow_info(self) -> dict:
        try:
            import tensorflow as tf

            return {
                "version": tf.version.VERSION,
                "git_version": tf.version.GIT_VERSION,
                "support_gpu": tf.test.is_built_with_cuda(),
            }
        except ImportError as error:
            if not self.silent:
                self.logger.exception(error)
            return None
        except ModuleNotFoundError as error:
            if not self.silent:
                self.logger.exception(error)
            return None


def parse_arguments():
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--silent",
        required=False,
        action="store_true",
        help="Do not print error message",
    )
    parser.set_defaults(silent=False)

    args = parser.parse_args()
    return args


def get_machine_info(silent=True) -> str:
    machine = MachineInfo(silent)
    return json.dumps(machine.machine_info, indent=2)


def get_device_info(silent=True) -> str:
    machine = MachineInfo(silent)
    info = machine.machine_info
    if info:
        info = {key: value for key, value in info.items() if key in ["gpu", "cpu", "memory"]}
    return json.dumps(info, indent=2)


if __name__ == "__main__":
    args = parse_arguments()
    print(get_machine_info(args.silent))

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\io\orc.py ===
""" orc compat """
from __future__ import annotations

import io
from types import ModuleType
from typing import (
    TYPE_CHECKING,
    Any,
    Literal,
)

from pandas._libs import lib
from pandas.compat._optional import import_optional_dependency
from pandas.util._validators import check_dtype_backend

from pandas.core.indexes.api import default_index

from pandas.io._util import arrow_table_to_pandas
from pandas.io.common import (
    get_handle,
    is_fsspec_url,
)

if TYPE_CHECKING:
    import fsspec
    import pyarrow.fs

    from pandas._typing import (
        DtypeBackend,
        FilePath,
        ReadBuffer,
        WriteBuffer,
    )

    from pandas.core.frame import DataFrame


def read_orc(
    path: FilePath | ReadBuffer[bytes],
    columns: list[str] | None = None,
    dtype_backend: DtypeBackend | lib.NoDefault = lib.no_default,
    filesystem: pyarrow.fs.FileSystem | fsspec.spec.AbstractFileSystem | None = None,
    **kwargs: Any,
) -> DataFrame:
    """
    Load an ORC object from the file path, returning a DataFrame.

    Parameters
    ----------
    path : str, path object, or file-like object
        String, path object (implementing ``os.PathLike[str]``), or file-like
        object implementing a binary ``read()`` function. The string could be a URL.
        Valid URL schemes include http, ftp, s3, and file. For file URLs, a host is
        expected. A local file could be:
        ``file://localhost/path/to/table.orc``.
    columns : list, default None
        If not None, only these columns will be read from the file.
        Output always follows the ordering of the file and not the columns list.
        This mirrors the original behaviour of
        :external+pyarrow:py:meth:`pyarrow.orc.ORCFile.read`.
    dtype_backend : {'numpy_nullable', 'pyarrow'}, default 'numpy_nullable'
        Back-end data type applied to the resultant :class:`DataFrame`
        (still experimental). Behaviour is as follows:

        * ``"numpy_nullable"``: returns nullable-dtype-backed :class:`DataFrame`
          (default).
        * ``"pyarrow"``: returns pyarrow-backed nullable :class:`ArrowDtype`
          DataFrame.

        .. versionadded:: 2.0

    filesystem : fsspec or pyarrow filesystem, default None
        Filesystem object to use when reading the parquet file.

        .. versionadded:: 2.1.0

    **kwargs
        Any additional kwargs are passed to pyarrow.

    Returns
    -------
    DataFrame

    Notes
    -----
    Before using this function you should read the :ref:`user guide about ORC <io.orc>`
    and :ref:`install optional dependencies <install.warn_orc>`.

    If ``path`` is a URI scheme pointing to a local or remote file (e.g. "s3://"),
    a ``pyarrow.fs`` filesystem will be attempted to read the file. You can also pass a
    pyarrow or fsspec filesystem object into the filesystem keyword to override this
    behavior.

    Examples
    --------
    >>> result = pd.read_orc("example_pa.orc")  # doctest: +SKIP
    """
    # we require a newer version of pyarrow than we support for parquet

    orc = import_optional_dependency("pyarrow.orc")

    check_dtype_backend(dtype_backend)

    with get_handle(path, "rb", is_text=False) as handles:
        source = handles.handle
        if is_fsspec_url(path) and filesystem is None:
            pa = import_optional_dependency("pyarrow")
            pa_fs = import_optional_dependency("pyarrow.fs")
            try:
                filesystem, source = pa_fs.FileSystem.from_uri(path)
            except (TypeError, pa.ArrowInvalid):
                pass

        pa_table = orc.read_table(
            source=source, columns=columns, filesystem=filesystem, **kwargs
        )
    return arrow_table_to_pandas(pa_table, dtype_backend=dtype_backend)


def to_orc(
    df: DataFrame,
    path: FilePath | WriteBuffer[bytes] | None = None,
    *,
    engine: Literal["pyarrow"] = "pyarrow",
    index: bool | None = None,
    engine_kwargs: dict[str, Any] | None = None,
) -> bytes | None:
    """
    Write a DataFrame to the ORC format.

    .. versionadded:: 1.5.0

    Parameters
    ----------
    df : DataFrame
        The dataframe to be written to ORC. Raises NotImplementedError
        if dtype of one or more columns is category, unsigned integers,
        intervals, periods or sparse.
    path : str, file-like object or None, default None
        If a string, it will be used as Root Directory path
        when writing a partitioned dataset. By file-like object,
        we refer to objects with a write() method, such as a file handle
        (e.g. via builtin open function). If path is None,
        a bytes object is returned.
    engine : str, default 'pyarrow'
        ORC library to use.
    index : bool, optional
        If ``True``, include the dataframe's index(es) in the file output. If
        ``False``, they will not be written to the file.
        If ``None``, similar to ``infer`` the dataframe's index(es)
        will be saved. However, instead of being saved as values,
        the RangeIndex will be stored as a range in the metadata so it
        doesn't require much space and is faster. Other indexes will
        be included as columns in the file output.
    engine_kwargs : dict[str, Any] or None, default None
        Additional keyword arguments passed to :func:`pyarrow.orc.write_table`.

    Returns
    -------
    bytes if no path argument is provided else None

    Raises
    ------
    NotImplementedError
        Dtype of one or more columns is category, unsigned integers, interval,
        period or sparse.
    ValueError
        engine is not pyarrow.

    Notes
    -----
    * Before using this function you should read the
      :ref:`user guide about ORC <io.orc>` and
      :ref:`install optional dependencies <install.warn_orc>`.
    * This function requires `pyarrow <https://arrow.apache.org/docs/python/>`_
      library.
    * For supported dtypes please refer to `supported ORC features in Arrow
      <https://arrow.apache.org/docs/cpp/orc.html#data-types>`__.
    * Currently timezones in datetime columns are not preserved when a
      dataframe is converted into ORC files.
    """
    if index is None:
        index = df.index.names[0] is not None
    if engine_kwargs is None:
        engine_kwargs = {}

    # validate index
    # --------------

    # validate that we have only a default index
    # raise on anything else as we don't serialize the index

    if not df.index.equals(default_index(len(df))):
        raise ValueError(
            "orc does not support serializing a non-default index for the index; "
            "you can .reset_index() to make the index into column(s)"
        )

    if df.index.name is not None:
        raise ValueError("orc does not serialize index meta-data on a default index")

    if engine != "pyarrow":
        raise ValueError("engine must be 'pyarrow'")
    engine = import_optional_dependency(engine, min_version="10.0.1")
    pa = import_optional_dependency("pyarrow")
    orc = import_optional_dependency("pyarrow.orc")

    was_none = path is None
    if was_none:
        path = io.BytesIO()
    assert path is not None  # For mypy
    with get_handle(path, "wb", is_text=False) as handles:
        assert isinstance(engine, ModuleType)  # For mypy
        try:
            orc.write_table(
                engine.Table.from_pandas(df, preserve_index=index),
                handles.handle,
                **engine_kwargs,
            )
        except (TypeError, pa.ArrowNotImplementedError) as e:
            raise NotImplementedError(
                "The dtype of one or more columns is not supported yet."
            ) from e

    if was_none:
        assert isinstance(path, io.BytesIO)  # For mypy
        return path.getvalue()
    return None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\commands\cache.py ===
import os
import textwrap
from optparse import Values
from typing import Any, List

from pip._internal.cli.base_command import Command
from pip._internal.cli.status_codes import ERROR, SUCCESS
from pip._internal.exceptions import CommandError, PipError
from pip._internal.utils import filesystem
from pip._internal.utils.logging import getLogger
from pip._internal.utils.misc import format_size

logger = getLogger(__name__)


class CacheCommand(Command):
    """
    Inspect and manage pip's wheel cache.

    Subcommands:

    - dir: Show the cache directory.
    - info: Show information about the cache.
    - list: List filenames of packages stored in the cache.
    - remove: Remove one or more package from the cache.
    - purge: Remove all items from the cache.

    ``<pattern>`` can be a glob expression or a package name.
    """

    ignore_require_venv = True
    usage = """
        %prog dir
        %prog info
        %prog list [<pattern>] [--format=[human, abspath]]
        %prog remove <pattern>
        %prog purge
    """

    def add_options(self) -> None:
        self.cmd_opts.add_option(
            "--format",
            action="store",
            dest="list_format",
            default="human",
            choices=("human", "abspath"),
            help="Select the output format among: human (default) or abspath",
        )

        self.parser.insert_option_group(0, self.cmd_opts)

    def run(self, options: Values, args: List[str]) -> int:
        handlers = {
            "dir": self.get_cache_dir,
            "info": self.get_cache_info,
            "list": self.list_cache_items,
            "remove": self.remove_cache_items,
            "purge": self.purge_cache,
        }

        if not options.cache_dir:
            logger.error("pip cache commands can not function since cache is disabled.")
            return ERROR

        # Determine action
        if not args or args[0] not in handlers:
            logger.error(
                "Need an action (%s) to perform.",
                ", ".join(sorted(handlers)),
            )
            return ERROR

        action = args[0]

        # Error handling happens here, not in the action-handlers.
        try:
            handlers[action](options, args[1:])
        except PipError as e:
            logger.error(e.args[0])
            return ERROR

        return SUCCESS

    def get_cache_dir(self, options: Values, args: List[Any]) -> None:
        if args:
            raise CommandError("Too many arguments")

        logger.info(options.cache_dir)

    def get_cache_info(self, options: Values, args: List[Any]) -> None:
        if args:
            raise CommandError("Too many arguments")

        num_http_files = len(self._find_http_files(options))
        num_packages = len(self._find_wheels(options, "*"))

        http_cache_location = self._cache_dir(options, "http-v2")
        old_http_cache_location = self._cache_dir(options, "http")
        wheels_cache_location = self._cache_dir(options, "wheels")
        http_cache_size = filesystem.format_size(
            filesystem.directory_size(http_cache_location)
            + filesystem.directory_size(old_http_cache_location)
        )
        wheels_cache_size = filesystem.format_directory_size(wheels_cache_location)

        message = (
            textwrap.dedent(
                """
                    Package index page cache location (pip v23.3+): {http_cache_location}
                    Package index page cache location (older pips): {old_http_cache_location}
                    Package index page cache size: {http_cache_size}
                    Number of HTTP files: {num_http_files}
                    Locally built wheels location: {wheels_cache_location}
                    Locally built wheels size: {wheels_cache_size}
                    Number of locally built wheels: {package_count}
                """  # noqa: E501
            )
            .format(
                http_cache_location=http_cache_location,
                old_http_cache_location=old_http_cache_location,
                http_cache_size=http_cache_size,
                num_http_files=num_http_files,
                wheels_cache_location=wheels_cache_location,
                package_count=num_packages,
                wheels_cache_size=wheels_cache_size,
            )
            .strip()
        )

        logger.info(message)

    def list_cache_items(self, options: Values, args: List[Any]) -> None:
        if len(args) > 1:
            raise CommandError("Too many arguments")

        if args:
            pattern = args[0]
        else:
            pattern = "*"

        files = self._find_wheels(options, pattern)
        if options.list_format == "human":
            self.format_for_human(files)
        else:
            self.format_for_abspath(files)

    def format_for_human(self, files: List[str]) -> None:
        if not files:
            logger.info("No locally built wheels cached.")
            return

        results = []
        for filename in files:
            wheel = os.path.basename(filename)
            size = filesystem.format_file_size(filename)
            results.append(f" - {wheel} ({size})")
        logger.info("Cache contents:\n")
        logger.info("\n".join(sorted(results)))

    def format_for_abspath(self, files: List[str]) -> None:
        if files:
            logger.info("\n".join(sorted(files)))

    def remove_cache_items(self, options: Values, args: List[Any]) -> None:
        if len(args) > 1:
            raise CommandError("Too many arguments")

        if not args:
            raise CommandError("Please provide a pattern")

        files = self._find_wheels(options, args[0])

        no_matching_msg = "No matching packages"
        if args[0] == "*":
            # Only fetch http files if no specific pattern given
            files += self._find_http_files(options)
        else:
            # Add the pattern to the log message
            no_matching_msg += f' for pattern "{args[0]}"'

        if not files:
            logger.warning(no_matching_msg)

        bytes_removed = 0
        for filename in files:
            bytes_removed += os.stat(filename).st_size
            os.unlink(filename)
            logger.verbose("Removed %s", filename)
        logger.info("Files removed: %s (%s)", len(files), format_size(bytes_removed))

    def purge_cache(self, options: Values, args: List[Any]) -> None:
        if args:
            raise CommandError("Too many arguments")

        return self.remove_cache_items(options, ["*"])

    def _cache_dir(self, options: Values, subdir: str) -> str:
        return os.path.join(options.cache_dir, subdir)

    def _find_http_files(self, options: Values) -> List[str]:
        old_http_dir = self._cache_dir(options, "http")
        new_http_dir = self._cache_dir(options, "http-v2")
        return filesystem.find_files(old_http_dir, "*") + filesystem.find_files(
            new_http_dir, "*"
        )

    def _find_wheels(self, options: Values, pattern: str) -> List[str]:
        wheel_dir = self._cache_dir(options, "wheels")

        # The wheel filename format, as specified in PEP 427, is:
        #     {distribution}-{version}(-{build})?-{python}-{abi}-{platform}.whl
        #
        # Additionally, non-alphanumeric values in the distribution are
        # normalized to underscores (_), meaning hyphens can never occur
        # before `-{version}`.
        #
        # Given that information:
        # - If the pattern we're given contains a hyphen (-), the user is
        #   providing at least the version. Thus, we can just append `*.whl`
        #   to match the rest of it.
        # - If the pattern we're given doesn't contain a hyphen (-), the
        #   user is only providing the name. Thus, we append `-*.whl` to
        #   match the hyphen before the version, followed by anything else.
        #
        # PEP 427: https://www.python.org/dev/peps/pep-0427/
        pattern = pattern + ("*.whl" if "-" in pattern else "-*.whl")

        return filesystem.find_files(wheel_dir, pattern)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\commands\show.py ===
import logging
import string
from optparse import Values
from typing import Generator, Iterable, Iterator, List, NamedTuple, Optional

from pip._vendor.packaging.requirements import InvalidRequirement
from pip._vendor.packaging.utils import canonicalize_name

from pip._internal.cli.base_command import Command
from pip._internal.cli.status_codes import ERROR, SUCCESS
from pip._internal.metadata import BaseDistribution, get_default_environment
from pip._internal.utils.misc import write_output

logger = logging.getLogger(__name__)


def normalize_project_url_label(label: str) -> str:
    # This logic is from PEP 753 (Well-known Project URLs in Metadata).
    chars_to_remove = string.punctuation + string.whitespace
    removal_map = str.maketrans("", "", chars_to_remove)
    return label.translate(removal_map).lower()


class ShowCommand(Command):
    """
    Show information about one or more installed packages.

    The output is in RFC-compliant mail header format.
    """

    usage = """
      %prog [options] <package> ..."""
    ignore_require_venv = True

    def add_options(self) -> None:
        self.cmd_opts.add_option(
            "-f",
            "--files",
            dest="files",
            action="store_true",
            default=False,
            help="Show the full list of installed files for each package.",
        )

        self.parser.insert_option_group(0, self.cmd_opts)

    def run(self, options: Values, args: List[str]) -> int:
        if not args:
            logger.warning("ERROR: Please provide a package name or names.")
            return ERROR
        query = args

        results = search_packages_info(query)
        if not print_results(
            results, list_files=options.files, verbose=options.verbose
        ):
            return ERROR
        return SUCCESS


class _PackageInfo(NamedTuple):
    name: str
    version: str
    location: str
    editable_project_location: Optional[str]
    requires: List[str]
    required_by: List[str]
    installer: str
    metadata_version: str
    classifiers: List[str]
    summary: str
    homepage: str
    project_urls: List[str]
    author: str
    author_email: str
    license: str
    license_expression: str
    entry_points: List[str]
    files: Optional[List[str]]


def search_packages_info(query: List[str]) -> Generator[_PackageInfo, None, None]:
    """
    Gather details from installed distributions. Print distribution name,
    version, location, and installed files. Installed files requires a
    pip generated 'installed-files.txt' in the distributions '.egg-info'
    directory.
    """
    env = get_default_environment()

    installed = {dist.canonical_name: dist for dist in env.iter_all_distributions()}
    query_names = [canonicalize_name(name) for name in query]
    missing = sorted(
        [name for name, pkg in zip(query, query_names) if pkg not in installed]
    )
    if missing:
        logger.warning("Package(s) not found: %s", ", ".join(missing))

    def _get_requiring_packages(current_dist: BaseDistribution) -> Iterator[str]:
        return (
            dist.metadata["Name"] or "UNKNOWN"
            for dist in installed.values()
            if current_dist.canonical_name
            in {canonicalize_name(d.name) for d in dist.iter_dependencies()}
        )

    for query_name in query_names:
        try:
            dist = installed[query_name]
        except KeyError:
            continue

        try:
            requires = sorted(
                # Avoid duplicates in requirements (e.g. due to environment markers).
                {req.name for req in dist.iter_dependencies()},
                key=str.lower,
            )
        except InvalidRequirement:
            requires = sorted(dist.iter_raw_dependencies(), key=str.lower)

        try:
            required_by = sorted(_get_requiring_packages(dist), key=str.lower)
        except InvalidRequirement:
            required_by = ["#N/A"]

        try:
            entry_points_text = dist.read_text("entry_points.txt")
            entry_points = entry_points_text.splitlines(keepends=False)
        except FileNotFoundError:
            entry_points = []

        files_iter = dist.iter_declared_entries()
        if files_iter is None:
            files: Optional[List[str]] = None
        else:
            files = sorted(files_iter)

        metadata = dist.metadata

        project_urls = metadata.get_all("Project-URL", [])
        homepage = metadata.get("Home-page", "")
        if not homepage:
            # It's common that there is a "homepage" Project-URL, but Home-page
            # remains unset (especially as PEP 621 doesn't surface the field).
            for url in project_urls:
                url_label, url = url.split(",", maxsplit=1)
                normalized_label = normalize_project_url_label(url_label)
                if normalized_label == "homepage":
                    homepage = url.strip()
                    break

        yield _PackageInfo(
            name=dist.raw_name,
            version=dist.raw_version,
            location=dist.location or "",
            editable_project_location=dist.editable_project_location,
            requires=requires,
            required_by=required_by,
            installer=dist.installer,
            metadata_version=dist.metadata_version or "",
            classifiers=metadata.get_all("Classifier", []),
            summary=metadata.get("Summary", ""),
            homepage=homepage,
            project_urls=project_urls,
            author=metadata.get("Author", ""),
            author_email=metadata.get("Author-email", ""),
            license=metadata.get("License", ""),
            license_expression=metadata.get("License-Expression", ""),
            entry_points=entry_points,
            files=files,
        )


def print_results(
    distributions: Iterable[_PackageInfo],
    list_files: bool,
    verbose: bool,
) -> bool:
    """
    Print the information from installed distributions found.
    """
    results_printed = False
    for i, dist in enumerate(distributions):
        results_printed = True
        if i > 0:
            write_output("---")

        metadata_version_tuple = tuple(map(int, dist.metadata_version.split(".")))

        write_output("Name: %s", dist.name)
        write_output("Version: %s", dist.version)
        write_output("Summary: %s", dist.summary)
        write_output("Home-page: %s", dist.homepage)
        write_output("Author: %s", dist.author)
        write_output("Author-email: %s", dist.author_email)
        if metadata_version_tuple >= (2, 4) and dist.license_expression:
            write_output("License-Expression: %s", dist.license_expression)
        else:
            write_output("License: %s", dist.license)
        write_output("Location: %s", dist.location)
        if dist.editable_project_location is not None:
            write_output(
                "Editable project location: %s", dist.editable_project_location
            )
        write_output("Requires: %s", ", ".join(dist.requires))
        write_output("Required-by: %s", ", ".join(dist.required_by))

        if verbose:
            write_output("Metadata-Version: %s", dist.metadata_version)
            write_output("Installer: %s", dist.installer)
            write_output("Classifiers:")
            for classifier in dist.classifiers:
                write_output("  %s", classifier)
            write_output("Entry-points:")
            for entry in dist.entry_points:
                write_output("  %s", entry.strip())
            write_output("Project-URLs:")
            for project_url in dist.project_urls:
                write_output("  %s", project_url)
        if list_files:
            write_output("Files:")
            if dist.files is None:
                write_output("Cannot locate RECORD or installed-files.txt")
            else:
                for line in dist.files:
                    write_output("  %s", line.strip())
    return results_printed

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\metadata\importlib\_dists.py ===
import email.message
import importlib.metadata
import pathlib
import zipfile
from os import PathLike
from typing import (
    Collection,
    Dict,
    Iterable,
    Iterator,
    Mapping,
    Optional,
    Sequence,
    Union,
    cast,
)

from pip._vendor.packaging.requirements import Requirement
from pip._vendor.packaging.utils import NormalizedName, canonicalize_name
from pip._vendor.packaging.version import Version
from pip._vendor.packaging.version import parse as parse_version

from pip._internal.exceptions import InvalidWheel, UnsupportedWheel
from pip._internal.metadata.base import (
    BaseDistribution,
    BaseEntryPoint,
    InfoPath,
    Wheel,
)
from pip._internal.utils.misc import normalize_path
from pip._internal.utils.packaging import get_requirement
from pip._internal.utils.temp_dir import TempDirectory
from pip._internal.utils.wheel import parse_wheel, read_wheel_metadata_file

from ._compat import (
    BasePath,
    get_dist_canonical_name,
    parse_name_and_version_from_info_directory,
)


class WheelDistribution(importlib.metadata.Distribution):
    """An ``importlib.metadata.Distribution`` read from a wheel.

    Although ``importlib.metadata.PathDistribution`` accepts ``zipfile.Path``,
    its implementation is too "lazy" for pip's needs (we can't keep the ZipFile
    handle open for the entire lifetime of the distribution object).

    This implementation eagerly reads the entire metadata directory into the
    memory instead, and operates from that.
    """

    def __init__(
        self,
        files: Mapping[pathlib.PurePosixPath, bytes],
        info_location: pathlib.PurePosixPath,
    ) -> None:
        self._files = files
        self.info_location = info_location

    @classmethod
    def from_zipfile(
        cls,
        zf: zipfile.ZipFile,
        name: str,
        location: str,
    ) -> "WheelDistribution":
        info_dir, _ = parse_wheel(zf, name)
        paths = (
            (name, pathlib.PurePosixPath(name.split("/", 1)[-1]))
            for name in zf.namelist()
            if name.startswith(f"{info_dir}/")
        )
        files = {
            relpath: read_wheel_metadata_file(zf, fullpath)
            for fullpath, relpath in paths
        }
        info_location = pathlib.PurePosixPath(location, info_dir)
        return cls(files, info_location)

    def iterdir(self, path: InfoPath) -> Iterator[pathlib.PurePosixPath]:
        # Only allow iterating through the metadata directory.
        if pathlib.PurePosixPath(str(path)) in self._files:
            return iter(self._files)
        raise FileNotFoundError(path)

    def read_text(self, filename: str) -> Optional[str]:
        try:
            data = self._files[pathlib.PurePosixPath(filename)]
        except KeyError:
            return None
        try:
            text = data.decode("utf-8")
        except UnicodeDecodeError as e:
            wheel = self.info_location.parent
            error = f"Error decoding metadata for {wheel}: {e} in {filename} file"
            raise UnsupportedWheel(error)
        return text

    def locate_file(self, path: Union[str, "PathLike[str]"]) -> pathlib.Path:
        # This method doesn't make sense for our in-memory wheel, but the API
        # requires us to define it.
        raise NotImplementedError


class Distribution(BaseDistribution):
    def __init__(
        self,
        dist: importlib.metadata.Distribution,
        info_location: Optional[BasePath],
        installed_location: Optional[BasePath],
    ) -> None:
        self._dist = dist
        self._info_location = info_location
        self._installed_location = installed_location

    @classmethod
    def from_directory(cls, directory: str) -> BaseDistribution:
        info_location = pathlib.Path(directory)
        dist = importlib.metadata.Distribution.at(info_location)
        return cls(dist, info_location, info_location.parent)

    @classmethod
    def from_metadata_file_contents(
        cls,
        metadata_contents: bytes,
        filename: str,
        project_name: str,
    ) -> BaseDistribution:
        # Generate temp dir to contain the metadata file, and write the file contents.
        temp_dir = pathlib.Path(
            TempDirectory(kind="metadata", globally_managed=True).path
        )
        metadata_path = temp_dir / "METADATA"
        metadata_path.write_bytes(metadata_contents)
        # Construct dist pointing to the newly created directory.
        dist = importlib.metadata.Distribution.at(metadata_path.parent)
        return cls(dist, metadata_path.parent, None)

    @classmethod
    def from_wheel(cls, wheel: Wheel, name: str) -> BaseDistribution:
        try:
            with wheel.as_zipfile() as zf:
                dist = WheelDistribution.from_zipfile(zf, name, wheel.location)
        except zipfile.BadZipFile as e:
            raise InvalidWheel(wheel.location, name) from e
        return cls(dist, dist.info_location, pathlib.PurePosixPath(wheel.location))

    @property
    def location(self) -> Optional[str]:
        if self._info_location is None:
            return None
        return str(self._info_location.parent)

    @property
    def info_location(self) -> Optional[str]:
        if self._info_location is None:
            return None
        return str(self._info_location)

    @property
    def installed_location(self) -> Optional[str]:
        if self._installed_location is None:
            return None
        return normalize_path(str(self._installed_location))

    @property
    def canonical_name(self) -> NormalizedName:
        return get_dist_canonical_name(self._dist)

    @property
    def version(self) -> Version:
        if version := parse_name_and_version_from_info_directory(self._dist)[1]:
            return parse_version(version)
        return parse_version(self._dist.version)

    @property
    def raw_version(self) -> str:
        return self._dist.version

    def is_file(self, path: InfoPath) -> bool:
        return self._dist.read_text(str(path)) is not None

    def iter_distutils_script_names(self) -> Iterator[str]:
        # A distutils installation is always "flat" (not in e.g. egg form), so
        # if this distribution's info location is NOT a pathlib.Path (but e.g.
        # zipfile.Path), it can never contain any distutils scripts.
        if not isinstance(self._info_location, pathlib.Path):
            return
        for child in self._info_location.joinpath("scripts").iterdir():
            yield child.name

    def read_text(self, path: InfoPath) -> str:
        content = self._dist.read_text(str(path))
        if content is None:
            raise FileNotFoundError(path)
        return content

    def iter_entry_points(self) -> Iterable[BaseEntryPoint]:
        # importlib.metadata's EntryPoint structure satisfies BaseEntryPoint.
        return self._dist.entry_points

    def _metadata_impl(self) -> email.message.Message:
        # From Python 3.10+, importlib.metadata declares PackageMetadata as the
        # return type. This protocol is unfortunately a disaster now and misses
        # a ton of fields that we need, including get() and get_payload(). We
        # rely on the implementation that the object is actually a Message now,
        # until upstream can improve the protocol. (python/cpython#94952)
        return cast(email.message.Message, self._dist.metadata)

    def iter_provided_extras(self) -> Iterable[NormalizedName]:
        return [
            canonicalize_name(extra)
            for extra in self.metadata.get_all("Provides-Extra", [])
        ]

    def iter_dependencies(self, extras: Collection[str] = ()) -> Iterable[Requirement]:
        contexts: Sequence[Dict[str, str]] = [{"extra": e} for e in extras]
        for req_string in self.metadata.get_all("Requires-Dist", []):
            # strip() because email.message.Message.get_all() may return a leading \n
            # in case a long header was wrapped.
            req = get_requirement(req_string.strip())
            if not req.marker:
                yield req
            elif not extras and req.marker.evaluate({"extra": ""}):
                yield req
            elif any(req.marker.evaluate(context) for context in contexts):
                yield req

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\urllib3\contrib\socks.py ===
"""
This module contains provisional support for SOCKS proxies from within
urllib3. This module supports SOCKS4, SOCKS4A (an extension of SOCKS4), and
SOCKS5. To enable its functionality, either install PySocks or install this
module with the ``socks`` extra.

The SOCKS implementation supports the full range of urllib3 features. It also
supports the following SOCKS features:

- SOCKS4A (``proxy_url='socks4a://...``)
- SOCKS4 (``proxy_url='socks4://...``)
- SOCKS5 with remote DNS (``proxy_url='socks5h://...``)
- SOCKS5 with local DNS (``proxy_url='socks5://...``)
- Usernames and passwords for the SOCKS proxy

.. note::
   It is recommended to use ``socks5h://`` or ``socks4a://`` schemes in
   your ``proxy_url`` to ensure that DNS resolution is done from the remote
   server instead of client-side when connecting to a domain name.

SOCKS4 supports IPv4 and domain names with the SOCKS4A extension. SOCKS5
supports IPv4, IPv6, and domain names.

When connecting to a SOCKS4 proxy the ``username`` portion of the ``proxy_url``
will be sent as the ``userid`` section of the SOCKS request:

.. code-block:: python

    proxy_url="socks4a://<userid>@proxy-host"

When connecting to a SOCKS5 proxy the ``username`` and ``password`` portion
of the ``proxy_url`` will be sent as the username/password to authenticate
with the proxy:

.. code-block:: python

    proxy_url="socks5h://<username>:<password>@proxy-host"

"""

from __future__ import annotations

try:
    import socks  # type: ignore[import-not-found]
except ImportError:
    import warnings

    from ..exceptions import DependencyWarning

    warnings.warn(
        (
            "SOCKS support in urllib3 requires the installation of optional "
            "dependencies: specifically, PySocks.  For more information, see "
            "https://urllib3.readthedocs.io/en/latest/advanced-usage.html#socks-proxies"
        ),
        DependencyWarning,
    )
    raise

import typing
from socket import timeout as SocketTimeout

from ..connection import HTTPConnection, HTTPSConnection
from ..connectionpool import HTTPConnectionPool, HTTPSConnectionPool
from ..exceptions import ConnectTimeoutError, NewConnectionError
from ..poolmanager import PoolManager
from ..util.url import parse_url

try:
    import ssl
except ImportError:
    ssl = None  # type: ignore[assignment]


class _TYPE_SOCKS_OPTIONS(typing.TypedDict):
    socks_version: int
    proxy_host: str | None
    proxy_port: str | None
    username: str | None
    password: str | None
    rdns: bool


class SOCKSConnection(HTTPConnection):
    """
    A plain-text HTTP connection that connects via a SOCKS proxy.
    """

    def __init__(
        self,
        _socks_options: _TYPE_SOCKS_OPTIONS,
        *args: typing.Any,
        **kwargs: typing.Any,
    ) -> None:
        self._socks_options = _socks_options
        super().__init__(*args, **kwargs)

    def _new_conn(self) -> socks.socksocket:
        """
        Establish a new connection via the SOCKS proxy.
        """
        extra_kw: dict[str, typing.Any] = {}
        if self.source_address:
            extra_kw["source_address"] = self.source_address

        if self.socket_options:
            extra_kw["socket_options"] = self.socket_options

        try:
            conn = socks.create_connection(
                (self.host, self.port),
                proxy_type=self._socks_options["socks_version"],
                proxy_addr=self._socks_options["proxy_host"],
                proxy_port=self._socks_options["proxy_port"],
                proxy_username=self._socks_options["username"],
                proxy_password=self._socks_options["password"],
                proxy_rdns=self._socks_options["rdns"],
                timeout=self.timeout,
                **extra_kw,
            )

        except SocketTimeout as e:
            raise ConnectTimeoutError(
                self,
                f"Connection to {self.host} timed out. (connect timeout={self.timeout})",
            ) from e

        except socks.ProxyError as e:
            # This is fragile as hell, but it seems to be the only way to raise
            # useful errors here.
            if e.socket_err:
                error = e.socket_err
                if isinstance(error, SocketTimeout):
                    raise ConnectTimeoutError(
                        self,
                        f"Connection to {self.host} timed out. (connect timeout={self.timeout})",
                    ) from e
                else:
                    # Adding `from e` messes with coverage somehow, so it's omitted.
                    # See #2386.
                    raise NewConnectionError(
                        self, f"Failed to establish a new connection: {error}"
                    )
            else:
                raise NewConnectionError(
                    self, f"Failed to establish a new connection: {e}"
                ) from e

        except OSError as e:  # Defensive: PySocks should catch all these.
            raise NewConnectionError(
                self, f"Failed to establish a new connection: {e}"
            ) from e

        return conn


# We don't need to duplicate the Verified/Unverified distinction from
# urllib3/connection.py here because the HTTPSConnection will already have been
# correctly set to either the Verified or Unverified form by that module. This
# means the SOCKSHTTPSConnection will automatically be the correct type.
class SOCKSHTTPSConnection(SOCKSConnection, HTTPSConnection):
    pass


class SOCKSHTTPConnectionPool(HTTPConnectionPool):
    ConnectionCls = SOCKSConnection


class SOCKSHTTPSConnectionPool(HTTPSConnectionPool):
    ConnectionCls = SOCKSHTTPSConnection


class SOCKSProxyManager(PoolManager):
    """
    A version of the urllib3 ProxyManager that routes connections via the
    defined SOCKS proxy.
    """

    pool_classes_by_scheme = {
        "http": SOCKSHTTPConnectionPool,
        "https": SOCKSHTTPSConnectionPool,
    }

    def __init__(
        self,
        proxy_url: str,
        username: str | None = None,
        password: str | None = None,
        num_pools: int = 10,
        headers: typing.Mapping[str, str] | None = None,
        **connection_pool_kw: typing.Any,
    ):
        parsed = parse_url(proxy_url)

        if username is None and password is None and parsed.auth is not None:
            split = parsed.auth.split(":")
            if len(split) == 2:
                username, password = split
        if parsed.scheme == "socks5":
            socks_version = socks.PROXY_TYPE_SOCKS5
            rdns = False
        elif parsed.scheme == "socks5h":
            socks_version = socks.PROXY_TYPE_SOCKS5
            rdns = True
        elif parsed.scheme == "socks4":
            socks_version = socks.PROXY_TYPE_SOCKS4
            rdns = False
        elif parsed.scheme == "socks4a":
            socks_version = socks.PROXY_TYPE_SOCKS4
            rdns = True
        else:
            raise ValueError(f"Unable to determine SOCKS version from {proxy_url}")

        self.proxy_url = proxy_url

        socks_options = {
            "socks_version": socks_version,
            "proxy_host": parsed.host,
            "proxy_port": parsed.port,
            "username": username,
            "password": password,
            "rdns": rdns,
        }
        connection_pool_kw["_socks_options"] = socks_options

        super().__init__(num_pools, headers, **connection_pool_kw)

        self.pool_classes_by_scheme = SOCKSProxyManager.pool_classes_by_scheme

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\testing\fixtures\orm.py ===
# testing/fixtures/orm.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors
from __future__ import annotations

from typing import Any

import sqlalchemy as sa
from .base import TestBase
from .sql import TablesTest
from .. import assertions
from .. import config
from .. import schema
from ..entities import BasicEntity
from ..entities import ComparableEntity
from ..util import adict
from ... import orm
from ...orm import DeclarativeBase
from ...orm import events as orm_events
from ...orm import registry


class ORMTest(TestBase):
    @config.fixture
    def fixture_session(self):
        return fixture_session()


class MappedTest(ORMTest, TablesTest, assertions.AssertsExecutionResults):
    # 'once', 'each', None
    run_setup_classes = "once"

    # 'once', 'each', None
    run_setup_mappers = "each"

    classes: Any = None

    @config.fixture(autouse=True, scope="class")
    def _setup_tables_test_class(self):
        cls = self.__class__
        cls._init_class()

        if cls.classes is None:
            cls.classes = adict()

        cls._setup_once_tables()
        cls._setup_once_classes()
        cls._setup_once_mappers()
        cls._setup_once_inserts()

        yield

        cls._teardown_once_class()
        cls._teardown_once_metadata_bind()

    @config.fixture(autouse=True, scope="function")
    def _setup_tables_test_instance(self):
        self._setup_each_tables()
        self._setup_each_classes()
        self._setup_each_mappers()
        self._setup_each_inserts()

        yield

        orm.session.close_all_sessions()
        self._teardown_each_mappers()
        self._teardown_each_classes()
        self._teardown_each_tables()

    @classmethod
    def _teardown_once_class(cls):
        cls.classes.clear()

    @classmethod
    def _setup_once_classes(cls):
        if cls.run_setup_classes == "once":
            cls._with_register_classes(cls.setup_classes)

    @classmethod
    def _setup_once_mappers(cls):
        if cls.run_setup_mappers == "once":
            cls.mapper_registry, cls.mapper = cls._generate_registry()
            cls._with_register_classes(cls.setup_mappers)

    def _setup_each_mappers(self):
        if self.run_setup_mappers != "once":
            (
                self.__class__.mapper_registry,
                self.__class__.mapper,
            ) = self._generate_registry()

        if self.run_setup_mappers == "each":
            self._with_register_classes(self.setup_mappers)

    def _setup_each_classes(self):
        if self.run_setup_classes == "each":
            self._with_register_classes(self.setup_classes)

    @classmethod
    def _generate_registry(cls):
        decl = registry(metadata=cls._tables_metadata)
        return decl, decl.map_imperatively

    @classmethod
    def _with_register_classes(cls, fn):
        """Run a setup method, framing the operation with a Base class
        that will catch new subclasses to be established within
        the "classes" registry.

        """
        cls_registry = cls.classes

        class _Base:
            def __init_subclass__(cls) -> None:
                assert cls_registry is not None
                cls_registry[cls.__name__] = cls
                super().__init_subclass__()

        class Basic(BasicEntity, _Base):
            pass

        class Comparable(ComparableEntity, _Base):
            pass

        cls.Basic = Basic
        cls.Comparable = Comparable
        fn()

    def _teardown_each_mappers(self):
        # some tests create mappers in the test bodies
        # and will define setup_mappers as None -
        # clear mappers in any case
        if self.run_setup_mappers != "once":
            orm.clear_mappers()

    def _teardown_each_classes(self):
        if self.run_setup_classes != "once":
            self.classes.clear()

    @classmethod
    def setup_classes(cls):
        pass

    @classmethod
    def setup_mappers(cls):
        pass


class DeclarativeMappedTest(MappedTest):
    run_setup_classes = "once"
    run_setup_mappers = "once"

    @classmethod
    def _setup_once_tables(cls):
        pass

    @classmethod
    def _with_register_classes(cls, fn):
        cls_registry = cls.classes

        class _DeclBase(DeclarativeBase):
            __table_cls__ = schema.Table
            metadata = cls._tables_metadata
            type_annotation_map = {
                str: sa.String().with_variant(
                    sa.String(50), "mysql", "mariadb", "oracle"
                )
            }

            def __init_subclass__(cls, **kw) -> None:
                assert cls_registry is not None
                cls_registry[cls.__name__] = cls
                super().__init_subclass__(**kw)

        cls.DeclarativeBasic = _DeclBase

        # sets up cls.Basic which is helpful for things like composite
        # classes
        super()._with_register_classes(fn)

        if cls._tables_metadata.tables and cls.run_create_tables:
            cls._tables_metadata.create_all(config.db)


class RemoveORMEventsGlobally:
    @config.fixture(autouse=True)
    def _remove_listeners(self):
        yield
        orm_events.MapperEvents._clear()
        orm_events.InstanceEvents._clear()
        orm_events.SessionEvents._clear()
        orm_events.InstrumentationEvents._clear()
        orm_events.QueryEvents._clear()


_fixture_sessions = set()


def fixture_session(**kw):
    kw.setdefault("autoflush", True)
    kw.setdefault("expire_on_commit", True)

    bind = kw.pop("bind", config.db)

    sess = orm.Session(bind, **kw)
    _fixture_sessions.add(sess)
    return sess


def close_all_sessions():
    # will close all still-referenced sessions
    orm.close_all_sessions()
    _fixture_sessions.clear()


def stop_test_class_inside_fixtures(cls):
    close_all_sessions()
    orm.clear_mappers()


def after_test():
    if _fixture_sessions:
        close_all_sessions()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\core\random.py ===
"""
When you need to use random numbers in SymPy library code, import from here
so there is only one generator working for SymPy. Imports from here should
behave the same as if they were being imported from Python's random module.
But only the routines currently used in SymPy are included here. To use others
import ``rng`` and access the method directly. For example, to capture the
current state of the generator use ``rng.getstate()``.

There is intentionally no Random to import from here. If you want
to control the state of the generator, import ``seed`` and call it
with or without an argument to set the state.

Examples
========

>>> from sympy.core.random import random, seed
>>> assert random() < 1
>>> seed(1); a = random()
>>> b = random()
>>> seed(1); c = random()
>>> assert a == c
>>> assert a != b  # remote possibility this will fail

"""
from sympy.utilities.iterables import is_sequence
from sympy.utilities.misc import as_int

import random as _random
rng = _random.Random()

choice = rng.choice
random = rng.random
randint = rng.randint
randrange = rng.randrange
sample = rng.sample
# seed = rng.seed
shuffle = rng.shuffle
uniform = rng.uniform

_assumptions_rng = _random.Random()
_assumptions_shuffle = _assumptions_rng.shuffle


def seed(a=None, version=2):
    rng.seed(a=a, version=version)
    _assumptions_rng.seed(a=a, version=version)


def random_complex_number(a=2, b=-1, c=3, d=1, rational=False, tolerance=None):
    """
    Return a random complex number.

    To reduce chance of hitting branch cuts or anything, we guarantee
    b <= Im z <= d, a <= Re z <= c

    When rational is True, a rational approximation to a random number
    is obtained within specified tolerance, if any.
    """
    from sympy.core.numbers import I
    from sympy.simplify.simplify import nsimplify
    A, B = uniform(a, c), uniform(b, d)
    if not rational:
        return A + I*B
    return (nsimplify(A, rational=True, tolerance=tolerance) +
        I*nsimplify(B, rational=True, tolerance=tolerance))


def verify_numerically(f, g, z=None, tol=1.0e-6, a=2, b=-1, c=3, d=1):
    """
    Test numerically that f and g agree when evaluated in the argument z.

    If z is None, all symbols will be tested. This routine does not test
    whether there are Floats present with precision higher than 15 digits
    so if there are, your results may not be what you expect due to round-
    off errors.

    Examples
    ========

    >>> from sympy import sin, cos
    >>> from sympy.abc import x
    >>> from sympy.core.random import verify_numerically as tn
    >>> tn(sin(x)**2 + cos(x)**2, 1, x)
    True
    """
    from sympy.core.symbol import Symbol
    from sympy.core.sympify import sympify
    from sympy.core.numbers import comp
    f, g = (sympify(i) for i in (f, g))
    if z is None:
        z = f.free_symbols | g.free_symbols
    elif isinstance(z, Symbol):
        z = [z]
    reps = list(zip(z, [random_complex_number(a, b, c, d) for _ in z]))
    z1 = f.subs(reps).n()
    z2 = g.subs(reps).n()
    return comp(z1, z2, tol)


def test_derivative_numerically(f, z, tol=1.0e-6, a=2, b=-1, c=3, d=1):
    """
    Test numerically that the symbolically computed derivative of f
    with respect to z is correct.

    This routine does not test whether there are Floats present with
    precision higher than 15 digits so if there are, your results may
    not be what you expect due to round-off errors.

    Examples
    ========

    >>> from sympy import sin
    >>> from sympy.abc import x
    >>> from sympy.core.random import test_derivative_numerically as td
    >>> td(sin(x), x)
    True
    """
    from sympy.core.numbers import comp
    from sympy.core.function import Derivative
    z0 = random_complex_number(a, b, c, d)
    f1 = f.diff(z).subs(z, z0)
    f2 = Derivative(f, z).doit_numerically(z0)
    return comp(f1.n(), f2.n(), tol)


def _randrange(seed=None):
    """Return a randrange generator.

    ``seed`` can be

    * None - return randomly seeded generator
    * int - return a generator seeded with the int
    * list - the values to be returned will be taken from the list
      in the order given; the provided list is not modified.

    Examples
    ========

    >>> from sympy.core.random import _randrange
    >>> rr = _randrange()
    >>> rr(1000) # doctest: +SKIP
    999
    >>> rr = _randrange(3)
    >>> rr(1000) # doctest: +SKIP
    238
    >>> rr = _randrange([0, 5, 1, 3, 4])
    >>> rr(3), rr(3)
    (0, 1)
    """
    if seed is None:
        return randrange
    elif isinstance(seed, int):
        rng.seed(seed)
        return randrange
    elif is_sequence(seed):
        seed = list(seed)  # make a copy
        seed.reverse()

        def give(a, b=None, seq=seed):
            if b is None:
                a, b = 0, a
            a, b = as_int(a), as_int(b)
            w = b - a
            if w < 1:
                raise ValueError('_randrange got empty range')
            try:
                x = seq.pop()
            except IndexError:
                raise ValueError('_randrange sequence was too short')
            if a <= x < b:
                return x
            else:
                return give(a, b, seq)
        return give
    else:
        raise ValueError('_randrange got an unexpected seed')


def _randint(seed=None):
    """Return a randint generator.

    ``seed`` can be

    * None - return randomly seeded generator
    * int - return a generator seeded with the int
    * list - the values to be returned will be taken from the list
      in the order given; the provided list is not modified.

    Examples
    ========

    >>> from sympy.core.random import _randint
    >>> ri = _randint()
    >>> ri(1, 1000) # doctest: +SKIP
    999
    >>> ri = _randint(3)
    >>> ri(1, 1000) # doctest: +SKIP
    238
    >>> ri = _randint([0, 5, 1, 2, 4])
    >>> ri(1, 3), ri(1, 3)
    (1, 2)
    """
    if seed is None:
        return randint
    elif isinstance(seed, int):
        rng.seed(seed)
        return randint
    elif is_sequence(seed):
        seed = list(seed)  # make a copy
        seed.reverse()

        def give(a, b, seq=seed):
            a, b = as_int(a), as_int(b)
            w = b - a
            if w < 0:
                raise ValueError('_randint got empty range')
            try:
                x = seq.pop()
            except IndexError:
                raise ValueError('_randint sequence was too short')
            if a <= x <= b:
                return x
            else:
                return give(a, b, seq)
        return give
    else:
        raise ValueError('_randint got an unexpected seed')

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\onnx_model_tnlr.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
import logging

from fusion_attention import AttentionMask, FusionAttention
from fusion_utils import NumpyHelper
from onnx import NodeProto, helper
from onnx_model import OnnxModel
from onnx_model_bert import BertOnnxModel

logger = logging.getLogger(__name__)


class FusionTnlrAttention(FusionAttention):
    """
    Fuse TNLR Attention subgraph into one Attention node.
    TNLR Attention has extra addition after qk nodes and adopts [S, B, NH] as I/O shape.
    """

    def __init__(
        self,
        model: OnnxModel,
        hidden_size: int,
        num_heads: int,
        attention_mask: AttentionMask,
    ):
        super().__init__(model, hidden_size, num_heads, attention_mask)

    def create_attention_node(
        self,
        mask_index: str,
        matmul: NodeProto,
        add: NodeProto,
        num_heads: int,
        hidden_size: int,
        input: str,
        output: str,
        add_qk_str: str,
    ) -> NodeProto | None:
        assert num_heads > 0
        if hidden_size > 0 and (hidden_size % num_heads) != 0:
            logger.debug(f"input hidden size {hidden_size} is not a multiple of num of heads {num_heads}")
            return None

        weight = self.model.get_initializer(matmul.input[1])
        bias = self.model.get_initializer(add.input[1]) or self.model.get_initializer(add.input[0])

        if weight is None or bias is None:
            return None

        qkv_weight = NumpyHelper.to_array(weight)
        qkv_bias = NumpyHelper.to_array(bias)

        attention_node_name = self.model.create_node_name("Attention")

        tensor_dtype = weight.data_type
        np_type = helper.tensor_dtype_to_np_dtype(tensor_dtype)
        weight = helper.make_tensor(
            name=attention_node_name + "_qkv_weight",
            data_type=tensor_dtype,
            dims=[hidden_size, 3 * hidden_size],
            vals=qkv_weight.astype(np_type).tobytes(),
            raw=True,
        )
        self.model.add_initializer(weight, self.this_graph_name)

        bias = helper.make_tensor(
            name=attention_node_name + "_qkv_bias",
            data_type=tensor_dtype,
            dims=[3 * hidden_size],
            vals=qkv_bias.astype(np_type).tobytes(),
            raw=True,
        )
        self.model.add_initializer(bias, self.this_graph_name)

        attention_inputs = [
            input,
            attention_node_name + "_qkv_weight",
            attention_node_name + "_qkv_bias",
        ]
        if mask_index is not None:
            attention_inputs.append(mask_index)
        else:
            attention_inputs.append("")

        if add_qk_str is not None:
            attention_inputs.append("")
            attention_inputs.append(add_qk_str)

        attention_node = helper.make_node(
            "Attention",
            inputs=attention_inputs,
            outputs=[output],
            name=attention_node_name,
        )
        attention_node.domain = "com.microsoft"
        attention_node.attribute.extend([helper.make_attribute("num_heads", num_heads)])

        return attention_node

    def fuse(self, normalize_node, input_name_to_nodes, output_name_to_node):
        # Sometimes we can not fuse skiplayernormalization since the add before layernorm has an output that used by nodes outside skiplayernorm
        # Conceptually we treat add before layernorm as skiplayernorm node since they share the same pattern
        start_node = normalize_node
        if normalize_node.op_type != "SkipLayerNormalization":
            return

        # SkipLayerNormalization has two inputs, and one of them is the root input for attention.
        qkv_nodes = self.model.match_parent_path(
            start_node,
            ["Where", "Add", "MatMul", "Reshape", "Transpose", "MatMul"],
            [1, 1, 1, 0, 0, 0],
        )
        if qkv_nodes is not None:
            (_, _, matmul_below, reshape_qkv, transpose_qkv, matmul_qkv) = qkv_nodes
        else:
            return

        other_inputs = []
        for _i, input in enumerate(start_node.input):
            if input not in output_name_to_node:
                continue

            if input == qkv_nodes[0].output[0]:
                continue
            other_inputs.append(input)
        if len(other_inputs) != 1:
            return

        root_input = other_inputs[0]

        v_nodes = self.model.match_parent_path(
            matmul_qkv,
            ["Transpose", "Reshape", "Slice", "Add", "MatMul"],
            [1, 0, 0, 0, 1],
        )
        if v_nodes is None:
            return
        (_, _, _, add, matmul) = v_nodes

        upper_nodes = self.model.match_parent_path(matmul, ["Transpose"], [0])
        transpose = upper_nodes[0]

        qk_nodes = self.model.match_parent_path(matmul_qkv, ["Softmax", "Add", "MatMul"], [0, 0, 0])
        if qk_nodes is None:
            return
        (_, add_qk, matmul_qk) = qk_nodes

        q_nodes = self.model.match_parent_path(
            matmul_qk,
            ["Mul", "Transpose", "Reshape", "Slice", "Add", "MatMul"],
            [0, 0, 0, 0, 0, 1],
        )
        if q_nodes is None:
            return
        add = q_nodes[-2]
        matmul = q_nodes[-1]

        k_nodes = self.model.match_parent_path(
            matmul_qk,
            ["Transpose", "Reshape", "Slice", "Add", "MatMul"],
            [1, 0, 0, 0, 1],
        )
        if k_nodes is None:
            return
        add = k_nodes[-2]
        matmul = k_nodes[-1]

        relative_position_bias_nodes = self.model.match_parent_path(add_qk, ["Reshape", "Where"], [1, 0])
        if relative_position_bias_nodes is None:
            return

        if matmul.input[0] == root_input:
            mask_index = None
            attention_last_node = reshape_qkv
            # number of heads are same for all the paths, hence to create attention node, we pass the q_num_heads
            # the input_hidden_size represents the input hidden size, this is used as needed but hidden sizes for Q, K are extracted appropriately
            new_node = self.create_attention_node(
                mask_index,
                matmul,
                add,
                self.num_heads,
                self.hidden_size,
                root_input,
                attention_last_node.output[0],
                relative_position_bias_nodes[0].input[0],
            )
            if new_node is None:
                return

            self.nodes_to_add.append(new_node)
            self.node_name_to_graph_name[new_node.name] = self.this_graph_name

            # Add a transpose node after the attention node
            back_transpose = helper.make_node(
                "Transpose",
                ["back_transpose_in_" + new_node.name],
                [new_node.output[0]],
                "back_transpose_" + new_node.name,
                perm=[1, 0, 2],
            )
            self.model.add_node(back_transpose, self.this_graph_name)
            new_node.input[0] = transpose.input[0]
            new_node.output[0] = "back_transpose_in_" + new_node.name

            self.nodes_to_remove.extend([attention_last_node, transpose_qkv, matmul_qkv])
            self.nodes_to_remove.extend(qk_nodes)
            self.nodes_to_remove.extend(q_nodes)
            self.nodes_to_remove.extend(k_nodes)
            self.nodes_to_remove.extend(v_nodes)

            # Use prune graph to remove mask nodes since they are shared by all attention nodes.
            # self.nodes_to_remove.extend(mask_nodes)
            self.prune_graph = True


class TnlrOnnxModel(BertOnnxModel):
    def __init__(self, model, num_heads, hidden_size):
        super().__init__(model, num_heads, hidden_size)
        self.attention_mask = AttentionMask(self)
        self.attention_fusion = FusionTnlrAttention(self, self.hidden_size, self.num_heads, self.attention_mask)

    def fuse_attention(self):
        self.attention_fusion.apply()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pandas\core\array_algos\quantile.py ===
from __future__ import annotations

from typing import TYPE_CHECKING

import numpy as np

from pandas.core.dtypes.missing import (
    isna,
    na_value_for_dtype,
)

if TYPE_CHECKING:
    from pandas._typing import (
        ArrayLike,
        Scalar,
        npt,
    )


def quantile_compat(
    values: ArrayLike, qs: npt.NDArray[np.float64], interpolation: str
) -> ArrayLike:
    """
    Compute the quantiles of the given values for each quantile in `qs`.

    Parameters
    ----------
    values : np.ndarray or ExtensionArray
    qs : np.ndarray[float64]
    interpolation : str

    Returns
    -------
    np.ndarray or ExtensionArray
    """
    if isinstance(values, np.ndarray):
        fill_value = na_value_for_dtype(values.dtype, compat=False)
        mask = isna(values)
        return quantile_with_mask(values, mask, fill_value, qs, interpolation)
    else:
        return values._quantile(qs, interpolation)


def quantile_with_mask(
    values: np.ndarray,
    mask: npt.NDArray[np.bool_],
    fill_value,
    qs: npt.NDArray[np.float64],
    interpolation: str,
) -> np.ndarray:
    """
    Compute the quantiles of the given values for each quantile in `qs`.

    Parameters
    ----------
    values : np.ndarray
        For ExtensionArray, this is _values_for_factorize()[0]
    mask : np.ndarray[bool]
        mask = isna(values)
        For ExtensionArray, this is computed before calling _value_for_factorize
    fill_value : Scalar
        The value to interpret fill NA entries with
        For ExtensionArray, this is _values_for_factorize()[1]
    qs : np.ndarray[float64]
    interpolation : str
        Type of interpolation

    Returns
    -------
    np.ndarray

    Notes
    -----
    Assumes values is already 2D.  For ExtensionArray this means np.atleast_2d
    has been called on _values_for_factorize()[0]

    Quantile is computed along axis=1.
    """
    assert values.shape == mask.shape
    if values.ndim == 1:
        # unsqueeze, operate, re-squeeze
        values = np.atleast_2d(values)
        mask = np.atleast_2d(mask)
        res_values = quantile_with_mask(values, mask, fill_value, qs, interpolation)
        return res_values[0]

    assert values.ndim == 2

    is_empty = values.shape[1] == 0

    if is_empty:
        # create the array of na_values
        # 2d len(values) * len(qs)
        flat = np.array([fill_value] * len(qs))
        result = np.repeat(flat, len(values)).reshape(len(values), len(qs))
    else:
        result = _nanpercentile(
            values,
            qs * 100.0,
            na_value=fill_value,
            mask=mask,
            interpolation=interpolation,
        )

        result = np.asarray(result)
        result = result.T

    return result


def _nanpercentile_1d(
    values: np.ndarray,
    mask: npt.NDArray[np.bool_],
    qs: npt.NDArray[np.float64],
    na_value: Scalar,
    interpolation: str,
) -> Scalar | np.ndarray:
    """
    Wrapper for np.percentile that skips missing values, specialized to
    1-dimensional case.

    Parameters
    ----------
    values : array over which to find quantiles
    mask : ndarray[bool]
        locations in values that should be considered missing
    qs : np.ndarray[float64] of quantile indices to find
    na_value : scalar
        value to return for empty or all-null values
    interpolation : str

    Returns
    -------
    quantiles : scalar or array
    """
    # mask is Union[ExtensionArray, ndarray]
    values = values[~mask]

    if len(values) == 0:
        # Can't pass dtype=values.dtype here bc we might have na_value=np.nan
        #  with values.dtype=int64 see test_quantile_empty
        # equiv: 'np.array([na_value] * len(qs))' but much faster
        return np.full(len(qs), na_value)

    return np.percentile(
        values,
        qs,
        # error: No overload variant of "percentile" matches argument
        # types "ndarray[Any, Any]", "ndarray[Any, dtype[floating[_64Bit]]]"
        # , "Dict[str, str]"  [call-overload]
        method=interpolation,  # type: ignore[call-overload]
    )


def _nanpercentile(
    values: np.ndarray,
    qs: npt.NDArray[np.float64],
    *,
    na_value,
    mask: npt.NDArray[np.bool_],
    interpolation: str,
):
    """
    Wrapper for np.percentile that skips missing values.

    Parameters
    ----------
    values : np.ndarray[ndim=2]  over which to find quantiles
    qs : np.ndarray[float64] of quantile indices to find
    na_value : scalar
        value to return for empty or all-null values
    mask : np.ndarray[bool]
        locations in values that should be considered missing
    interpolation : str

    Returns
    -------
    quantiles : scalar or array
    """

    if values.dtype.kind in "mM":
        # need to cast to integer to avoid rounding errors in numpy
        result = _nanpercentile(
            values.view("i8"),
            qs=qs,
            na_value=na_value.view("i8"),
            mask=mask,
            interpolation=interpolation,
        )

        # Note: we have to do `astype` and not view because in general we
        #  have float result at this point, not i8
        return result.astype(values.dtype)

    if mask.any():
        # Caller is responsible for ensuring mask shape match
        assert mask.shape == values.shape
        result = [
            _nanpercentile_1d(val, m, qs, na_value, interpolation=interpolation)
            for (val, m) in zip(list(values), list(mask))
        ]
        if values.dtype.kind == "f":
            # preserve itemsize
            result = np.asarray(result, dtype=values.dtype).T
        else:
            result = np.asarray(result).T
            if (
                result.dtype != values.dtype
                and not mask.all()
                and (result == result.astype(values.dtype, copy=False)).all()
            ):
                # mask.all() will never get cast back to int
                # e.g. values id integer dtype and result is floating dtype,
                #  only cast back to integer dtype if result values are all-integer.
                result = result.astype(values.dtype, copy=False)
        return result
    else:
        return np.percentile(
            values,
            qs,
            axis=1,
            # error: No overload variant of "percentile" matches argument types
            # "ndarray[Any, Any]", "ndarray[Any, dtype[floating[_64Bit]]]",
            # "int", "Dict[str, str]"  [call-overload]
            method=interpolation,  # type: ignore[call-overload]
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\firefox\firefox_binary.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.


import os
import time
from platform import system
from subprocess import DEVNULL, STDOUT, Popen

from typing_extensions import deprecated

from selenium.common.exceptions import WebDriverException
from selenium.webdriver.common import utils


@deprecated("Use binary_location property in Firefox Options to set location")
class FirefoxBinary:
    NO_FOCUS_LIBRARY_NAME = "x_ignore_nofocus.so"

    def __init__(self, firefox_path=None, log_file=None):
        """Creates a new instance of Firefox binary.

        :Args:
         - firefox_path - Path to the Firefox executable. By default, it will be detected from the standard locations.
         - log_file - A file object to redirect the firefox process output to. It can be sys.stdout.
                      Please note that with parallel run the output won't be synchronous.
                      By default, it will be redirected to /dev/null.
        """
        self._start_cmd = firefox_path
        # We used to default to subprocess.PIPE instead of /dev/null, but after
        # a while the pipe would fill up and Firefox would freeze.
        self._log_file = log_file or DEVNULL
        self.command_line = None
        self.platform = system().lower()
        if not self._start_cmd:
            self._start_cmd = self._get_firefox_start_cmd()
        if not self._start_cmd.strip():
            raise WebDriverException(
                "Failed to find firefox binary. You can set it by specifying "
                "the path to 'firefox_binary':\n\nfrom "
                "selenium.webdriver.firefox.firefox_binary import "
                "FirefoxBinary\n\nbinary = "
                "FirefoxBinary('/path/to/binary')\ndriver = "
                "webdriver.Firefox(firefox_binary=binary)"
            )
        # Rather than modifying the environment of the calling Python process
        # copy it and modify as needed.
        self._firefox_env = os.environ.copy()
        self._firefox_env["MOZ_CRASHREPORTER_DISABLE"] = "1"
        self._firefox_env["MOZ_NO_REMOTE"] = "1"
        self._firefox_env["NO_EM_RESTART"] = "1"

    def add_command_line_options(self, *args):
        self.command_line = args

    def launch_browser(self, profile, timeout=30):
        """Launches the browser for the given profile name.

        It is assumed the profile already exists.
        """
        self.profile = profile

        self._start_from_profile_path(self.profile.path)
        self._wait_until_connectable(timeout=timeout)

    def kill(self):
        """Kill the browser.

        This is useful when the browser is stuck.
        """
        if self.process:
            self.process.kill()
            self.process.wait()

    def _start_from_profile_path(self, path):
        self._firefox_env["XRE_PROFILE_PATH"] = path

        if self.platform == "linux":
            self._modify_link_library_path()
        command = [self._start_cmd, "-foreground"]
        if self.command_line:
            for cli in self.command_line:
                command.append(cli)
        self.process = Popen(command, stdout=self._log_file, stderr=STDOUT, env=self._firefox_env)

    def _wait_until_connectable(self, timeout=30):
        """Blocks until the extension is connectable in the firefox."""
        count = 0
        while not utils.is_connectable(self.profile.port):
            if self.process.poll():
                # Browser has exited
                raise WebDriverException(
                    "The browser appears to have exited "
                    "before we could connect. If you specified a log_file in "
                    "the FirefoxBinary constructor, check it for details."
                )
            if count >= timeout:
                self.kill()
                raise WebDriverException(
                    "Can't load the profile. Possible firefox version mismatch. "
                    "You must use GeckoDriver instead for Firefox 48+. Profile "
                    f"Dir: {self.profile.path} If you specified a log_file in the "
                    "FirefoxBinary constructor, check it for details."
                )
            count += 1
            time.sleep(1)
        return True

    def _find_exe_in_registry(self):
        try:
            from _winreg import HKEY_CURRENT_USER, HKEY_LOCAL_MACHINE, OpenKey, QueryValue
        except ImportError:
            from winreg import HKEY_CURRENT_USER, HKEY_LOCAL_MACHINE, OpenKey, QueryValue
        import shlex

        keys = (
            r"SOFTWARE\Classes\FirefoxHTML\shell\open\command",
            r"SOFTWARE\Classes\Applications\firefox.exe\shell\open\command",
        )
        command = ""
        for path in keys:
            try:
                key = OpenKey(HKEY_LOCAL_MACHINE, path)
                command = QueryValue(key, "")
                break
            except OSError:
                try:
                    key = OpenKey(HKEY_CURRENT_USER, path)
                    command = QueryValue(key, "")
                    break
                except OSError:
                    pass
        else:
            return ""

        if not command:
            return ""

        return shlex.split(command)[0]

    def _get_firefox_start_cmd(self):
        """Return the command to start firefox."""
        start_cmd = ""
        if self.platform == "darwin":  # small darwin due to lower() in self.platform
            ffname = "firefox"
            start_cmd = self.which(ffname)
            # use hardcoded path if nothing else was found by which()
            if not start_cmd:
                start_cmd = "/Applications/Firefox.app/Contents/MacOS/firefox"
            # fallback to homebrew installation for mac users
            if not os.path.exists(start_cmd):
                start_cmd = os.path.expanduser("~") + start_cmd
        elif self.platform == "windows":  # same
            start_cmd = self._find_exe_in_registry() or self._default_windows_location()
        elif self.platform == "java" and os.name == "nt":
            start_cmd = self._default_windows_location()
        else:
            for ffname in ["firefox", "iceweasel"]:
                start_cmd = self.which(ffname)
                if start_cmd:
                    break
            else:
                # couldn't find firefox on the system path
                raise RuntimeError(
                    "Could not find firefox in your system PATH."
                    " Please specify the firefox binary location or install firefox"
                )
        return start_cmd

    def _default_windows_location(self):
        program_files = [
            os.getenv("PROGRAMFILES", r"C:\Program Files"),
            os.getenv("PROGRAMFILES(X86)", r"C:\Program Files (x86)"),
        ]
        for path in program_files:
            binary_path = os.path.join(path, r"Mozilla Firefox\firefox.exe")
            if os.access(binary_path, os.X_OK):
                return binary_path
        return ""

    def _modify_link_library_path(self):
        existing_ld_lib_path = os.environ.get("LD_LIBRARY_PATH", "")

        new_ld_lib_path = self._extract_and_check(self.profile, "x86", "amd64")

        new_ld_lib_path += existing_ld_lib_path

        self._firefox_env["LD_LIBRARY_PATH"] = new_ld_lib_path
        self._firefox_env["LD_PRELOAD"] = self.NO_FOCUS_LIBRARY_NAME

    def _extract_and_check(self, profile, x86, amd64):
        paths = [x86, amd64]
        built_path = ""
        for path in paths:
            library_path = os.path.join(profile.path, path)
            if not os.path.exists(library_path):
                os.makedirs(library_path)
            import shutil

            shutil.copy(os.path.join(os.path.dirname(__file__), path, self.NO_FOCUS_LIBRARY_NAME), library_path)
            built_path += library_path + ":"

        return built_path

    def which(self, fname):
        """Returns the fully qualified path by searching Path of the given
        name."""
        for pe in os.environ["PATH"].split(os.pathsep):
            checkname = os.path.join(pe, fname)
            if os.access(checkname, os.X_OK) and not os.path.isdir(checkname):
                return checkname
        return None

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\control.py ===
import sys
import time
from typing import TYPE_CHECKING, Callable, Dict, Iterable, List, Union

if sys.version_info >= (3, 8):
    from typing import Final
else:
    from pip._vendor.typing_extensions import Final  # pragma: no cover

from .segment import ControlCode, ControlType, Segment

if TYPE_CHECKING:
    from .console import Console, ConsoleOptions, RenderResult

STRIP_CONTROL_CODES: Final = [
    7,  # Bell
    8,  # Backspace
    11,  # Vertical tab
    12,  # Form feed
    13,  # Carriage return
]
_CONTROL_STRIP_TRANSLATE: Final = {
    _codepoint: None for _codepoint in STRIP_CONTROL_CODES
}

CONTROL_ESCAPE: Final = {
    7: "\\a",
    8: "\\b",
    11: "\\v",
    12: "\\f",
    13: "\\r",
}

CONTROL_CODES_FORMAT: Dict[int, Callable[..., str]] = {
    ControlType.BELL: lambda: "\x07",
    ControlType.CARRIAGE_RETURN: lambda: "\r",
    ControlType.HOME: lambda: "\x1b[H",
    ControlType.CLEAR: lambda: "\x1b[2J",
    ControlType.ENABLE_ALT_SCREEN: lambda: "\x1b[?1049h",
    ControlType.DISABLE_ALT_SCREEN: lambda: "\x1b[?1049l",
    ControlType.SHOW_CURSOR: lambda: "\x1b[?25h",
    ControlType.HIDE_CURSOR: lambda: "\x1b[?25l",
    ControlType.CURSOR_UP: lambda param: f"\x1b[{param}A",
    ControlType.CURSOR_DOWN: lambda param: f"\x1b[{param}B",
    ControlType.CURSOR_FORWARD: lambda param: f"\x1b[{param}C",
    ControlType.CURSOR_BACKWARD: lambda param: f"\x1b[{param}D",
    ControlType.CURSOR_MOVE_TO_COLUMN: lambda param: f"\x1b[{param+1}G",
    ControlType.ERASE_IN_LINE: lambda param: f"\x1b[{param}K",
    ControlType.CURSOR_MOVE_TO: lambda x, y: f"\x1b[{y+1};{x+1}H",
    ControlType.SET_WINDOW_TITLE: lambda title: f"\x1b]0;{title}\x07",
}


class Control:
    """A renderable that inserts a control code (non printable but may move cursor).

    Args:
        *codes (str): Positional arguments are either a :class:`~rich.segment.ControlType` enum or a
            tuple of ControlType and an integer parameter
    """

    __slots__ = ["segment"]

    def __init__(self, *codes: Union[ControlType, ControlCode]) -> None:
        control_codes: List[ControlCode] = [
            (code,) if isinstance(code, ControlType) else code for code in codes
        ]
        _format_map = CONTROL_CODES_FORMAT
        rendered_codes = "".join(
            _format_map[code](*parameters) for code, *parameters in control_codes
        )
        self.segment = Segment(rendered_codes, None, control_codes)

    @classmethod
    def bell(cls) -> "Control":
        """Ring the 'bell'."""
        return cls(ControlType.BELL)

    @classmethod
    def home(cls) -> "Control":
        """Move cursor to 'home' position."""
        return cls(ControlType.HOME)

    @classmethod
    def move(cls, x: int = 0, y: int = 0) -> "Control":
        """Move cursor relative to current position.

        Args:
            x (int): X offset.
            y (int): Y offset.

        Returns:
            ~Control: Control object.

        """

        def get_codes() -> Iterable[ControlCode]:
            control = ControlType
            if x:
                yield (
                    control.CURSOR_FORWARD if x > 0 else control.CURSOR_BACKWARD,
                    abs(x),
                )
            if y:
                yield (
                    control.CURSOR_DOWN if y > 0 else control.CURSOR_UP,
                    abs(y),
                )

        control = cls(*get_codes())
        return control

    @classmethod
    def move_to_column(cls, x: int, y: int = 0) -> "Control":
        """Move to the given column, optionally add offset to row.

        Returns:
            x (int): absolute x (column)
            y (int): optional y offset (row)

        Returns:
            ~Control: Control object.
        """

        return (
            cls(
                (ControlType.CURSOR_MOVE_TO_COLUMN, x),
                (
                    ControlType.CURSOR_DOWN if y > 0 else ControlType.CURSOR_UP,
                    abs(y),
                ),
            )
            if y
            else cls((ControlType.CURSOR_MOVE_TO_COLUMN, x))
        )

    @classmethod
    def move_to(cls, x: int, y: int) -> "Control":
        """Move cursor to absolute position.

        Args:
            x (int): x offset (column)
            y (int): y offset (row)

        Returns:
            ~Control: Control object.
        """
        return cls((ControlType.CURSOR_MOVE_TO, x, y))

    @classmethod
    def clear(cls) -> "Control":
        """Clear the screen."""
        return cls(ControlType.CLEAR)

    @classmethod
    def show_cursor(cls, show: bool) -> "Control":
        """Show or hide the cursor."""
        return cls(ControlType.SHOW_CURSOR if show else ControlType.HIDE_CURSOR)

    @classmethod
    def alt_screen(cls, enable: bool) -> "Control":
        """Enable or disable alt screen."""
        if enable:
            return cls(ControlType.ENABLE_ALT_SCREEN, ControlType.HOME)
        else:
            return cls(ControlType.DISABLE_ALT_SCREEN)

    @classmethod
    def title(cls, title: str) -> "Control":
        """Set the terminal window title

        Args:
            title (str): The new terminal window title
        """
        return cls((ControlType.SET_WINDOW_TITLE, title))

    def __str__(self) -> str:
        return self.segment.text

    def __rich_console__(
        self, console: "Console", options: "ConsoleOptions"
    ) -> "RenderResult":
        if self.segment.text:
            yield self.segment


def strip_control_codes(
    text: str, _translate_table: Dict[int, None] = _CONTROL_STRIP_TRANSLATE
) -> str:
    """Remove control codes from text.

    Args:
        text (str): A string possibly contain control codes.

    Returns:
        str: String with control codes removed.
    """
    return text.translate(_translate_table)


def escape_control_codes(
    text: str,
    _translate_table: Dict[int, str] = CONTROL_ESCAPE,
) -> str:
    """Replace control codes with their "escaped" equivalent in the given text.
    (e.g. "\b" becomes "\\b")

    Args:
        text (str): A string possibly containing control codes.

    Returns:
        str: String with control codes replaced with their escaped version.
    """
    return text.translate(_translate_table)


if __name__ == "__main__":  # pragma: no cover
    from pip._vendor.rich.console import Console

    console = Console()
    console.print("Look at the title of your terminal window ^")
    # console.print(Control((ControlType.SET_WINDOW_TITLE, "Hello, world!")))
    for i in range(10):
        console.set_window_title("🚀 Loading" + "." * i)
        time.sleep(0.5)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\dialects\mysql\dml.py ===
# dialects/mysql/dml.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
from __future__ import annotations

from typing import Any
from typing import Dict
from typing import List
from typing import Mapping
from typing import Optional
from typing import Tuple
from typing import Union

from ... import exc
from ... import util
from ...sql._typing import _DMLTableArgument
from ...sql.base import _exclusive_against
from ...sql.base import _generative
from ...sql.base import ColumnCollection
from ...sql.base import ReadOnlyColumnCollection
from ...sql.dml import Insert as StandardInsert
from ...sql.elements import ClauseElement
from ...sql.elements import KeyedColumnElement
from ...sql.expression import alias
from ...sql.selectable import NamedFromClause
from ...util.typing import Self


__all__ = ("Insert", "insert")


def insert(table: _DMLTableArgument) -> Insert:
    """Construct a MySQL/MariaDB-specific variant :class:`_mysql.Insert`
    construct.

    .. container:: inherited_member

        The :func:`sqlalchemy.dialects.mysql.insert` function creates
        a :class:`sqlalchemy.dialects.mysql.Insert`.  This class is based
        on the dialect-agnostic :class:`_sql.Insert` construct which may
        be constructed using the :func:`_sql.insert` function in
        SQLAlchemy Core.

    The :class:`_mysql.Insert` construct includes additional methods
    :meth:`_mysql.Insert.on_duplicate_key_update`.

    """
    return Insert(table)


class Insert(StandardInsert):
    """MySQL-specific implementation of INSERT.

    Adds methods for MySQL-specific syntaxes such as ON DUPLICATE KEY UPDATE.

    The :class:`~.mysql.Insert` object is created using the
    :func:`sqlalchemy.dialects.mysql.insert` function.

    .. versionadded:: 1.2

    """

    stringify_dialect = "mysql"
    inherit_cache = False

    @property
    def inserted(
        self,
    ) -> ReadOnlyColumnCollection[str, KeyedColumnElement[Any]]:
        """Provide the "inserted" namespace for an ON DUPLICATE KEY UPDATE
        statement

        MySQL's ON DUPLICATE KEY UPDATE clause allows reference to the row
        that would be inserted, via a special function called ``VALUES()``.
        This attribute provides all columns in this row to be referenceable
        such that they will render within a ``VALUES()`` function inside the
        ON DUPLICATE KEY UPDATE clause.    The attribute is named ``.inserted``
        so as not to conflict with the existing
        :meth:`_expression.Insert.values` method.

        .. tip::  The :attr:`_mysql.Insert.inserted` attribute is an instance
            of :class:`_expression.ColumnCollection`, which provides an
            interface the same as that of the :attr:`_schema.Table.c`
            collection described at :ref:`metadata_tables_and_columns`.
            With this collection, ordinary names are accessible like attributes
            (e.g. ``stmt.inserted.some_column``), but special names and
            dictionary method names should be accessed using indexed access,
            such as ``stmt.inserted["column name"]`` or
            ``stmt.inserted["values"]``.  See the docstring for
            :class:`_expression.ColumnCollection` for further examples.

        .. seealso::

            :ref:`mysql_insert_on_duplicate_key_update` - example of how
            to use :attr:`_expression.Insert.inserted`

        """
        return self.inserted_alias.columns

    @util.memoized_property
    def inserted_alias(self) -> NamedFromClause:
        return alias(self.table, name="inserted")

    @_generative
    @_exclusive_against(
        "_post_values_clause",
        msgs={
            "_post_values_clause": "This Insert construct already "
            "has an ON DUPLICATE KEY clause present"
        },
    )
    def on_duplicate_key_update(self, *args: _UpdateArg, **kw: Any) -> Self:
        r"""
        Specifies the ON DUPLICATE KEY UPDATE clause.

        :param \**kw:  Column keys linked to UPDATE values.  The
         values may be any SQL expression or supported literal Python
         values.

        .. warning:: This dictionary does **not** take into account
           Python-specified default UPDATE values or generation functions,
           e.g. those specified using :paramref:`_schema.Column.onupdate`.
           These values will not be exercised for an ON DUPLICATE KEY UPDATE
           style of UPDATE, unless values are manually specified here.

        :param \*args: As an alternative to passing key/value parameters,
         a dictionary or list of 2-tuples can be passed as a single positional
         argument.

         Passing a single dictionary is equivalent to the keyword argument
         form::

            insert().on_duplicate_key_update({"name": "some name"})

         Passing a list of 2-tuples indicates that the parameter assignments
         in the UPDATE clause should be ordered as sent, in a manner similar
         to that described for the :class:`_expression.Update`
         construct overall
         in :ref:`tutorial_parameter_ordered_updates`::

            insert().on_duplicate_key_update(
                [
                    ("name", "some name"),
                    ("value", "some value"),
                ]
            )

         .. versionchanged:: 1.3 parameters can be specified as a dictionary
            or list of 2-tuples; the latter form provides for parameter
            ordering.


        .. versionadded:: 1.2

        .. seealso::

            :ref:`mysql_insert_on_duplicate_key_update`

        """
        if args and kw:
            raise exc.ArgumentError(
                "Can't pass kwargs and positional arguments simultaneously"
            )

        if args:
            if len(args) > 1:
                raise exc.ArgumentError(
                    "Only a single dictionary or list of tuples "
                    "is accepted positionally."
                )
            values = args[0]
        else:
            values = kw

        self._post_values_clause = OnDuplicateClause(
            self.inserted_alias, values
        )
        return self


class OnDuplicateClause(ClauseElement):
    __visit_name__ = "on_duplicate_key_update"

    _parameter_ordering: Optional[List[str]] = None

    update: Dict[str, Any]
    stringify_dialect = "mysql"

    def __init__(
        self, inserted_alias: NamedFromClause, update: _UpdateArg
    ) -> None:
        self.inserted_alias = inserted_alias

        # auto-detect that parameters should be ordered.   This is copied from
        # Update._proces_colparams(), however we don't look for a special flag
        # in this case since we are not disambiguating from other use cases as
        # we are in Update.values().
        if isinstance(update, list) and (
            update and isinstance(update[0], tuple)
        ):
            self._parameter_ordering = [key for key, value in update]
            update = dict(update)

        if isinstance(update, dict):
            if not update:
                raise ValueError(
                    "update parameter dictionary must not be empty"
                )
        elif isinstance(update, ColumnCollection):
            update = dict(update)
        else:
            raise ValueError(
                "update parameter must be a non-empty dictionary "
                "or a ColumnCollection such as the `.c.` collection "
                "of a Table object"
            )
        self.update = update


_UpdateArg = Union[
    Mapping[Any, Any], List[Tuple[str, Any]], ColumnCollection[Any, Any]
]

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_tests\test_wait_for_object.py ===
import os

import pytest

on_windows = os.name == "nt"
# Mark all the tests in this file as being windows-only
pytestmark = pytest.mark.skipif(not on_windows, reason="windows only")

import trio

from .. import _core, _timeouts
from .._core._tests.tutil import slow

if on_windows:
    from .._core._windows_cffi import Handle, ffi, kernel32
    from .._wait_for_object import WaitForMultipleObjects_sync, WaitForSingleObject


def test_WaitForMultipleObjects_sync() -> None:
    # This does a series of tests where we set/close the handle before
    # initiating the waiting for it.
    #
    # Note that closing the handle (not signaling) will cause the
    # *initiation* of a wait to return immediately. But closing a handle
    # that is already being waited on will not stop whatever is waiting
    # for it.

    # One handle
    handle1 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    kernel32.SetEvent(handle1)
    WaitForMultipleObjects_sync(handle1)
    kernel32.CloseHandle(handle1)
    print("test_WaitForMultipleObjects_sync one OK")

    # Two handles, signal first
    handle1 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    handle2 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    kernel32.SetEvent(handle1)
    WaitForMultipleObjects_sync(handle1, handle2)
    kernel32.CloseHandle(handle1)
    kernel32.CloseHandle(handle2)
    print("test_WaitForMultipleObjects_sync set first OK")

    # Two handles, signal second
    handle1 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    handle2 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    kernel32.SetEvent(handle2)
    WaitForMultipleObjects_sync(handle1, handle2)
    kernel32.CloseHandle(handle1)
    kernel32.CloseHandle(handle2)
    print("test_WaitForMultipleObjects_sync set second OK")

    # Two handles, close first
    handle1 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    handle2 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    kernel32.CloseHandle(handle1)
    with pytest.raises(OSError, match=r"^\[WinError 6\] The handle is invalid$"):
        WaitForMultipleObjects_sync(handle1, handle2)
    kernel32.CloseHandle(handle2)
    print("test_WaitForMultipleObjects_sync close first OK")

    # Two handles, close second
    handle1 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    handle2 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    kernel32.CloseHandle(handle2)
    with pytest.raises(OSError, match=r"^\[WinError 6\] The handle is invalid$"):
        WaitForMultipleObjects_sync(handle1, handle2)
    kernel32.CloseHandle(handle1)
    print("test_WaitForMultipleObjects_sync close second OK")


@slow
async def test_WaitForMultipleObjects_sync_slow() -> None:
    # This does a series of test in which the main thread sync-waits for
    # handles, while we spawn a thread to set the handles after a short while.

    TIMEOUT = 0.3

    # One handle
    handle1 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    t0 = _core.current_time()
    async with _core.open_nursery() as nursery:
        nursery.start_soon(
            trio.to_thread.run_sync,
            WaitForMultipleObjects_sync,
            handle1,
        )
        await _timeouts.sleep(TIMEOUT)
        # If we would comment the line below, the above thread will be stuck,
        # and Trio won't exit this scope
        kernel32.SetEvent(handle1)
    t1 = _core.current_time()
    assert TIMEOUT <= (t1 - t0) < 2.0 * TIMEOUT
    kernel32.CloseHandle(handle1)
    print("test_WaitForMultipleObjects_sync_slow one OK")

    # Two handles, signal first
    handle1 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    handle2 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    t0 = _core.current_time()
    async with _core.open_nursery() as nursery:
        nursery.start_soon(
            trio.to_thread.run_sync,
            WaitForMultipleObjects_sync,
            handle1,
            handle2,
        )
        await _timeouts.sleep(TIMEOUT)
        kernel32.SetEvent(handle1)
    t1 = _core.current_time()
    assert TIMEOUT <= (t1 - t0) < 2.0 * TIMEOUT
    kernel32.CloseHandle(handle1)
    kernel32.CloseHandle(handle2)
    print("test_WaitForMultipleObjects_sync_slow thread-set first OK")

    # Two handles, signal second
    handle1 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    handle2 = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    t0 = _core.current_time()
    async with _core.open_nursery() as nursery:
        nursery.start_soon(
            trio.to_thread.run_sync,
            WaitForMultipleObjects_sync,
            handle1,
            handle2,
        )
        await _timeouts.sleep(TIMEOUT)
        kernel32.SetEvent(handle2)
    t1 = _core.current_time()
    assert TIMEOUT <= (t1 - t0) < 2.0 * TIMEOUT
    kernel32.CloseHandle(handle1)
    kernel32.CloseHandle(handle2)
    print("test_WaitForMultipleObjects_sync_slow thread-set second OK")


async def test_WaitForSingleObject() -> None:
    # This does a series of test for setting/closing the handle before
    # initiating the wait.

    # Test already set
    handle = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    kernel32.SetEvent(handle)
    await WaitForSingleObject(handle)  # should return at once
    kernel32.CloseHandle(handle)
    print("test_WaitForSingleObject already set OK")

    # Test already set, as int
    handle = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    handle_int = int(ffi.cast("intptr_t", handle))
    kernel32.SetEvent(handle)
    await WaitForSingleObject(handle_int)  # should return at once
    kernel32.CloseHandle(handle)
    print("test_WaitForSingleObject already set OK")

    # Test already closed
    handle = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    kernel32.CloseHandle(handle)
    with pytest.raises(OSError, match=r"^\[WinError 6\] The handle is invalid$"):
        await WaitForSingleObject(handle)  # should return at once
    print("test_WaitForSingleObject already closed OK")

    # Not a handle
    with pytest.raises(TypeError):
        await WaitForSingleObject("not a handle")  # type: ignore[arg-type] # Wrong type
    # with pytest.raises(OSError):
    #     await WaitForSingleObject(99)  # If you're unlucky, it actually IS a handle :(
    print("test_WaitForSingleObject not a handle OK")


@slow
async def test_WaitForSingleObject_slow() -> None:
    # This does a series of test for setting the handle in another task,
    # and cancelling the wait task.

    # Set the timeout used in the tests. We test the waiting time against
    # the timeout with a certain margin.
    TIMEOUT = 0.3

    async def signal_soon_async(handle: Handle) -> None:
        await _timeouts.sleep(TIMEOUT)
        kernel32.SetEvent(handle)

    # Test handle is SET after TIMEOUT in separate coroutine

    handle = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    t0 = _core.current_time()

    async with _core.open_nursery() as nursery:
        nursery.start_soon(WaitForSingleObject, handle)
        nursery.start_soon(signal_soon_async, handle)

    kernel32.CloseHandle(handle)
    t1 = _core.current_time()
    assert TIMEOUT <= (t1 - t0) < 2.0 * TIMEOUT
    print("test_WaitForSingleObject_slow set from task OK")

    # Test handle is SET after TIMEOUT in separate coroutine, as int

    handle = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    handle_int = int(ffi.cast("intptr_t", handle))
    t0 = _core.current_time()

    async with _core.open_nursery() as nursery:
        nursery.start_soon(WaitForSingleObject, handle_int)
        nursery.start_soon(signal_soon_async, handle)

    kernel32.CloseHandle(handle)
    t1 = _core.current_time()
    assert TIMEOUT <= (t1 - t0) < 2.0 * TIMEOUT
    print("test_WaitForSingleObject_slow set from task as int OK")

    # Test handle is CLOSED after 1 sec - NOPE see comment above

    # Test cancellation

    handle = kernel32.CreateEventA(ffi.NULL, True, False, ffi.NULL)
    t0 = _core.current_time()

    with _timeouts.move_on_after(TIMEOUT):
        await WaitForSingleObject(handle)

    kernel32.CloseHandle(handle)
    t1 = _core.current_time()
    assert TIMEOUT <= (t1 - t0) < 2.0 * TIMEOUT
    print("test_WaitForSingleObject_slow cancellation OK")

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\openpyxl\styles\fills.py ===

# Copyright (c) 2010-2024 openpyxl

from openpyxl.descriptors import (
    Float,
    Set,
    Alias,
    NoneSet,
    Sequence,
    Integer,
    MinMax,
)
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.compat import safe_string

from .colors import ColorDescriptor, Color

from openpyxl.xml.functions import Element, localname
from openpyxl.xml.constants import SHEET_MAIN_NS


FILL_NONE = 'none'
FILL_SOLID = 'solid'
FILL_PATTERN_DARKDOWN = 'darkDown'
FILL_PATTERN_DARKGRAY = 'darkGray'
FILL_PATTERN_DARKGRID = 'darkGrid'
FILL_PATTERN_DARKHORIZONTAL = 'darkHorizontal'
FILL_PATTERN_DARKTRELLIS = 'darkTrellis'
FILL_PATTERN_DARKUP = 'darkUp'
FILL_PATTERN_DARKVERTICAL = 'darkVertical'
FILL_PATTERN_GRAY0625 = 'gray0625'
FILL_PATTERN_GRAY125 = 'gray125'
FILL_PATTERN_LIGHTDOWN = 'lightDown'
FILL_PATTERN_LIGHTGRAY = 'lightGray'
FILL_PATTERN_LIGHTGRID = 'lightGrid'
FILL_PATTERN_LIGHTHORIZONTAL = 'lightHorizontal'
FILL_PATTERN_LIGHTTRELLIS = 'lightTrellis'
FILL_PATTERN_LIGHTUP = 'lightUp'
FILL_PATTERN_LIGHTVERTICAL = 'lightVertical'
FILL_PATTERN_MEDIUMGRAY = 'mediumGray'

fills = (FILL_SOLID, FILL_PATTERN_DARKDOWN, FILL_PATTERN_DARKGRAY,
         FILL_PATTERN_DARKGRID, FILL_PATTERN_DARKHORIZONTAL, FILL_PATTERN_DARKTRELLIS,
         FILL_PATTERN_DARKUP, FILL_PATTERN_DARKVERTICAL, FILL_PATTERN_GRAY0625,
         FILL_PATTERN_GRAY125, FILL_PATTERN_LIGHTDOWN, FILL_PATTERN_LIGHTGRAY,
         FILL_PATTERN_LIGHTGRID, FILL_PATTERN_LIGHTHORIZONTAL,
         FILL_PATTERN_LIGHTTRELLIS, FILL_PATTERN_LIGHTUP, FILL_PATTERN_LIGHTVERTICAL,
         FILL_PATTERN_MEDIUMGRAY)


class Fill(Serialisable):

    """Base class"""

    tagname = "fill"

    @classmethod
    def from_tree(cls, el):
        children = [c for c in el]
        if not children:
            return
        child = children[0]
        if "patternFill" in child.tag:
            return PatternFill._from_tree(child)
        return super(Fill, GradientFill).from_tree(child)


class PatternFill(Fill):
    """Area fill patterns for use in styles.
    Caution: if you do not specify a fill_type, other attributes will have
    no effect !"""

    tagname = "patternFill"

    __elements__ = ('fgColor', 'bgColor')

    patternType = NoneSet(values=fills)
    fill_type = Alias("patternType")
    fgColor = ColorDescriptor()
    start_color = Alias("fgColor")
    bgColor = ColorDescriptor()
    end_color = Alias("bgColor")

    def __init__(self, patternType=None, fgColor=Color(), bgColor=Color(),
                 fill_type=None, start_color=None, end_color=None):
        if fill_type is not None:
            patternType = fill_type
        self.patternType = patternType
        if start_color is not None:
            fgColor = start_color
        self.fgColor = fgColor
        if end_color is not None:
            bgColor = end_color
        self.bgColor = bgColor

    @classmethod
    def _from_tree(cls, el):
        attrib = dict(el.attrib)
        for child in el:
            desc = localname(child)
            attrib[desc] = Color.from_tree(child)
        return cls(**attrib)


    def to_tree(self, tagname=None, idx=None):
        parent = Element("fill")
        el = Element(self.tagname)
        if self.patternType is not None:
            el.set('patternType', self.patternType)
        for c in self.__elements__:
            value = getattr(self, c)
            if value != Color():
                el.append(value.to_tree(c))
        parent.append(el)
        return parent


DEFAULT_EMPTY_FILL = PatternFill()
DEFAULT_GRAY_FILL = PatternFill(patternType='gray125')


class Stop(Serialisable):

    tagname = "stop"

    position = MinMax(min=0, max=1)
    color = ColorDescriptor()

    def __init__(self, color, position):
        self.position = position
        self.color = color


def _assign_position(values):
    """
    Automatically assign positions if a list of colours is provided.

    It is not permitted to mix colours and stops
    """
    n_values = len(values)
    n_stops = sum(isinstance(value, Stop) for value in values)

    if n_stops == 0:
        interval = 1
        if n_values > 2:
            interval = 1 / (n_values - 1)
        values = [Stop(value, i * interval)
                  for i, value in enumerate(values)]

    elif n_stops < n_values:
        raise ValueError('Cannot interpret mix of Stops and Colors in GradientFill')

    pos = set()
    for stop in values:
        if stop.position in pos:
            raise ValueError("Duplicate position {0}".format(stop.position))
        pos.add(stop.position)

    return values


class StopList(Sequence):

    expected_type = Stop

    def __set__(self, obj, values):
        values = _assign_position(values)
        super().__set__(obj, values)


class GradientFill(Fill):
    """Fill areas with gradient

    Two types of gradient fill are supported:

        - A type='linear' gradient interpolates colours between
          a set of specified Stops, across the length of an area.
          The gradient is left-to-right by default, but this
          orientation can be modified with the degree
          attribute.  A list of Colors can be provided instead
          and they will be positioned with equal distance between them.

        - A type='path' gradient applies a linear gradient from each
          edge of the area. Attributes top, right, bottom, left specify
          the extent of fill from the respective borders. Thus top="0.2"
          will fill the top 20% of the cell.

    """

    tagname = "gradientFill"

    type = Set(values=('linear', 'path'))
    fill_type = Alias("type")
    degree = Float()
    left = Float()
    right = Float()
    top = Float()
    bottom = Float()
    stop = StopList()


    def __init__(self, type="linear", degree=0, left=0, right=0, top=0,
                 bottom=0, stop=()):
        self.degree = degree
        self.left = left
        self.right = right
        self.top = top
        self.bottom = bottom
        self.stop = stop
        self.type = type


    def __iter__(self):
        for attr in self.__attrs__:
            value = getattr(self, attr)
            if value:
                yield attr, safe_string(value)


    def to_tree(self, tagname=None, namespace=None, idx=None):
        parent = Element("fill")
        el = super().to_tree()
        parent.append(el)
        return parent

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_internal\models\direct_url.py ===
"""PEP 610"""

import json
import re
import urllib.parse
from dataclasses import dataclass
from typing import Any, ClassVar, Dict, Iterable, Optional, Type, TypeVar, Union

__all__ = [
    "DirectUrl",
    "DirectUrlValidationError",
    "DirInfo",
    "ArchiveInfo",
    "VcsInfo",
]

T = TypeVar("T")

DIRECT_URL_METADATA_NAME = "direct_url.json"
ENV_VAR_RE = re.compile(r"^\$\{[A-Za-z0-9-_]+\}(:\$\{[A-Za-z0-9-_]+\})?$")


class DirectUrlValidationError(Exception):
    pass


def _get(
    d: Dict[str, Any], expected_type: Type[T], key: str, default: Optional[T] = None
) -> Optional[T]:
    """Get value from dictionary and verify expected type."""
    if key not in d:
        return default
    value = d[key]
    if not isinstance(value, expected_type):
        raise DirectUrlValidationError(
            f"{value!r} has unexpected type for {key} (expected {expected_type})"
        )
    return value


def _get_required(
    d: Dict[str, Any], expected_type: Type[T], key: str, default: Optional[T] = None
) -> T:
    value = _get(d, expected_type, key, default)
    if value is None:
        raise DirectUrlValidationError(f"{key} must have a value")
    return value


def _exactly_one_of(infos: Iterable[Optional["InfoType"]]) -> "InfoType":
    infos = [info for info in infos if info is not None]
    if not infos:
        raise DirectUrlValidationError(
            "missing one of archive_info, dir_info, vcs_info"
        )
    if len(infos) > 1:
        raise DirectUrlValidationError(
            "more than one of archive_info, dir_info, vcs_info"
        )
    assert infos[0] is not None
    return infos[0]


def _filter_none(**kwargs: Any) -> Dict[str, Any]:
    """Make dict excluding None values."""
    return {k: v for k, v in kwargs.items() if v is not None}


@dataclass
class VcsInfo:
    name: ClassVar = "vcs_info"

    vcs: str
    commit_id: str
    requested_revision: Optional[str] = None

    @classmethod
    def _from_dict(cls, d: Optional[Dict[str, Any]]) -> Optional["VcsInfo"]:
        if d is None:
            return None
        return cls(
            vcs=_get_required(d, str, "vcs"),
            commit_id=_get_required(d, str, "commit_id"),
            requested_revision=_get(d, str, "requested_revision"),
        )

    def _to_dict(self) -> Dict[str, Any]:
        return _filter_none(
            vcs=self.vcs,
            requested_revision=self.requested_revision,
            commit_id=self.commit_id,
        )


class ArchiveInfo:
    name = "archive_info"

    def __init__(
        self,
        hash: Optional[str] = None,
        hashes: Optional[Dict[str, str]] = None,
    ) -> None:
        # set hashes before hash, since the hash setter will further populate hashes
        self.hashes = hashes
        self.hash = hash

    @property
    def hash(self) -> Optional[str]:
        return self._hash

    @hash.setter
    def hash(self, value: Optional[str]) -> None:
        if value is not None:
            # Auto-populate the hashes key to upgrade to the new format automatically.
            # We don't back-populate the legacy hash key from hashes.
            try:
                hash_name, hash_value = value.split("=", 1)
            except ValueError:
                raise DirectUrlValidationError(
                    f"invalid archive_info.hash format: {value!r}"
                )
            if self.hashes is None:
                self.hashes = {hash_name: hash_value}
            elif hash_name not in self.hashes:
                self.hashes = self.hashes.copy()
                self.hashes[hash_name] = hash_value
        self._hash = value

    @classmethod
    def _from_dict(cls, d: Optional[Dict[str, Any]]) -> Optional["ArchiveInfo"]:
        if d is None:
            return None
        return cls(hash=_get(d, str, "hash"), hashes=_get(d, dict, "hashes"))

    def _to_dict(self) -> Dict[str, Any]:
        return _filter_none(hash=self.hash, hashes=self.hashes)


@dataclass
class DirInfo:
    name: ClassVar = "dir_info"

    editable: bool = False

    @classmethod
    def _from_dict(cls, d: Optional[Dict[str, Any]]) -> Optional["DirInfo"]:
        if d is None:
            return None
        return cls(editable=_get_required(d, bool, "editable", default=False))

    def _to_dict(self) -> Dict[str, Any]:
        return _filter_none(editable=self.editable or None)


InfoType = Union[ArchiveInfo, DirInfo, VcsInfo]


@dataclass
class DirectUrl:
    url: str
    info: InfoType
    subdirectory: Optional[str] = None

    def _remove_auth_from_netloc(self, netloc: str) -> str:
        if "@" not in netloc:
            return netloc
        user_pass, netloc_no_user_pass = netloc.split("@", 1)
        if (
            isinstance(self.info, VcsInfo)
            and self.info.vcs == "git"
            and user_pass == "git"
        ):
            return netloc
        if ENV_VAR_RE.match(user_pass):
            return netloc
        return netloc_no_user_pass

    @property
    def redacted_url(self) -> str:
        """url with user:password part removed unless it is formed with
        environment variables as specified in PEP 610, or it is ``git``
        in the case of a git URL.
        """
        purl = urllib.parse.urlsplit(self.url)
        netloc = self._remove_auth_from_netloc(purl.netloc)
        surl = urllib.parse.urlunsplit(
            (purl.scheme, netloc, purl.path, purl.query, purl.fragment)
        )
        return surl

    def validate(self) -> None:
        self.from_dict(self.to_dict())

    @classmethod
    def from_dict(cls, d: Dict[str, Any]) -> "DirectUrl":
        return DirectUrl(
            url=_get_required(d, str, "url"),
            subdirectory=_get(d, str, "subdirectory"),
            info=_exactly_one_of(
                [
                    ArchiveInfo._from_dict(_get(d, dict, "archive_info")),
                    DirInfo._from_dict(_get(d, dict, "dir_info")),
                    VcsInfo._from_dict(_get(d, dict, "vcs_info")),
                ]
            ),
        )

    def to_dict(self) -> Dict[str, Any]:
        res = _filter_none(
            url=self.redacted_url,
            subdirectory=self.subdirectory,
        )
        res[self.info.name] = self.info._to_dict()
        return res

    @classmethod
    def from_json(cls, s: str) -> "DirectUrl":
        return cls.from_dict(json.loads(s))

    def to_json(self) -> str:
        return json.dumps(self.to_dict(), sort_keys=True)

    def is_local_editable(self) -> bool:
        return isinstance(self.info, DirInfo) and self.info.editable

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\testing\schema.py ===
# testing/schema.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php
# mypy: ignore-errors

from __future__ import annotations

import sys

from . import config
from . import exclusions
from .. import event
from .. import schema
from .. import types as sqltypes
from ..orm import mapped_column as _orm_mapped_column
from ..util import OrderedDict

__all__ = ["Table", "Column"]

table_options = {}


def Table(*args, **kw) -> schema.Table:
    """A schema.Table wrapper/hook for dialect-specific tweaks."""

    test_opts = {k: kw.pop(k) for k in list(kw) if k.startswith("test_")}

    kw.update(table_options)

    if exclusions.against(config._current, "mysql"):
        if (
            "mysql_engine" not in kw
            and "mysql_type" not in kw
            and "autoload_with" not in kw
        ):
            if "test_needs_fk" in test_opts or "test_needs_acid" in test_opts:
                kw["mysql_engine"] = "InnoDB"
            else:
                # there are in fact test fixtures that rely upon MyISAM,
                # due to MySQL / MariaDB having poor FK behavior under innodb,
                # such as a self-referential table can't be deleted from at
                # once without attending to per-row dependencies.  We'd need to
                # add special steps to some fixtures if we want to not
                # explicitly state MyISAM here
                kw["mysql_engine"] = "MyISAM"
    elif exclusions.against(config._current, "mariadb"):
        if (
            "mariadb_engine" not in kw
            and "mariadb_type" not in kw
            and "autoload_with" not in kw
        ):
            if "test_needs_fk" in test_opts or "test_needs_acid" in test_opts:
                kw["mariadb_engine"] = "InnoDB"
            else:
                kw["mariadb_engine"] = "MyISAM"

    return schema.Table(*args, **kw)


def mapped_column(*args, **kw):
    """An orm.mapped_column wrapper/hook for dialect-specific tweaks."""

    return _schema_column(_orm_mapped_column, args, kw)


def Column(*args, **kw):
    """A schema.Column wrapper/hook for dialect-specific tweaks."""

    return _schema_column(schema.Column, args, kw)


def _schema_column(factory, args, kw):
    test_opts = {k: kw.pop(k) for k in list(kw) if k.startswith("test_")}

    if not config.requirements.foreign_key_ddl.enabled_for_config(config):
        args = [arg for arg in args if not isinstance(arg, schema.ForeignKey)]

    construct = factory(*args, **kw)

    if factory is schema.Column:
        col = construct
    else:
        col = construct.column

    if test_opts.get("test_needs_autoincrement", False) and kw.get(
        "primary_key", False
    ):
        if col.default is None and col.server_default is None:
            col.autoincrement = True

        # allow any test suite to pick up on this
        col.info["test_needs_autoincrement"] = True

        # hardcoded rule for oracle; this should
        # be moved out
        if exclusions.against(config._current, "oracle"):

            def add_seq(c, tbl):
                c._init_items(
                    schema.Sequence(
                        _truncate_name(
                            config.db.dialect, tbl.name + "_" + c.name + "_seq"
                        ),
                        optional=True,
                    )
                )

            event.listen(col, "after_parent_attach", add_seq, propagate=True)
    return construct


class eq_type_affinity:
    """Helper to compare types inside of datastructures based on affinity.

    E.g.::

        eq_(
            inspect(connection).get_columns("foo"),
            [
                {
                    "name": "id",
                    "type": testing.eq_type_affinity(sqltypes.INTEGER),
                    "nullable": False,
                    "default": None,
                    "autoincrement": False,
                },
                {
                    "name": "data",
                    "type": testing.eq_type_affinity(sqltypes.NullType),
                    "nullable": True,
                    "default": None,
                    "autoincrement": False,
                },
            ],
        )

    """

    def __init__(self, target):
        self.target = sqltypes.to_instance(target)

    def __eq__(self, other):
        return self.target._type_affinity is other._type_affinity

    def __ne__(self, other):
        return self.target._type_affinity is not other._type_affinity


class eq_compile_type:
    """similar to eq_type_affinity but uses compile"""

    def __init__(self, target):
        self.target = target

    def __eq__(self, other):
        return self.target == other.compile()

    def __ne__(self, other):
        return self.target != other.compile()


class eq_clause_element:
    """Helper to compare SQL structures based on compare()"""

    def __init__(self, target):
        self.target = target

    def __eq__(self, other):
        return self.target.compare(other)

    def __ne__(self, other):
        return not self.target.compare(other)


def _truncate_name(dialect, name):
    if len(name) > dialect.max_identifier_length:
        return (
            name[0 : max(dialect.max_identifier_length - 6, 0)]
            + "_"
            + hex(hash(name) % 64)[2:]
        )
    else:
        return name


def pep435_enum(name):
    # Implements PEP 435 in the minimal fashion needed by SQLAlchemy
    __members__ = OrderedDict()

    def __init__(self, name, value, alias=None):
        self.name = name
        self.value = value
        self.__members__[name] = self
        value_to_member[value] = self
        setattr(self.__class__, name, self)
        if alias:
            self.__members__[alias] = self
            setattr(self.__class__, alias, self)

    value_to_member = {}

    @classmethod
    def get(cls, value):
        return value_to_member[value]

    someenum = type(
        name,
        (object,),
        {"__members__": __members__, "__init__": __init__, "get": get},
    )

    # getframe() trick for pickling I don't understand courtesy
    # Python namedtuple()
    try:
        module = sys._getframe(1).f_globals.get("__name__", "__main__")
    except (AttributeError, ValueError):
        pass
    if module is not None:
        someenum.__module__ = module

    return someenum

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\physics\quantum\qasm.py ===
"""

qasm.py - Functions to parse a set of qasm commands into a SymPy Circuit.

Examples taken from Chuang's page: https://web.archive.org/web/20220120121541/https://www.media.mit.edu/quanta/qasm2circ/

The code returns a circuit and an associated list of labels.

>>> from sympy.physics.quantum.qasm import Qasm
>>> q = Qasm('qubit q0', 'qubit q1', 'h q0', 'cnot q0,q1')
>>> q.get_circuit()
CNOT(1,0)*H(1)

>>> q = Qasm('qubit q0', 'qubit q1', 'cnot q0,q1', 'cnot q1,q0', 'cnot q0,q1')
>>> q.get_circuit()
CNOT(1,0)*CNOT(0,1)*CNOT(1,0)
"""

__all__ = [
    'Qasm',
    ]

from math import prod

from sympy.physics.quantum.gate import H, CNOT, X, Z, CGate, CGateS, SWAP, S, T,CPHASE
from sympy.physics.quantum.circuitplot import Mz

def read_qasm(lines):
    return Qasm(*lines.splitlines())

def read_qasm_file(filename):
    return Qasm(*open(filename).readlines())

def flip_index(i, n):
    """Reorder qubit indices from largest to smallest.

    >>> from sympy.physics.quantum.qasm import flip_index
    >>> flip_index(0, 2)
    1
    >>> flip_index(1, 2)
    0
    """
    return n-i-1

def trim(line):
    """Remove everything following comment # characters in line.

    >>> from sympy.physics.quantum.qasm import trim
    >>> trim('nothing happens here')
    'nothing happens here'
    >>> trim('something #happens here')
    'something '
    """
    if '#' not in line:
        return line
    return line.split('#')[0]

def get_index(target, labels):
    """Get qubit labels from the rest of the line,and return indices

    >>> from sympy.physics.quantum.qasm import get_index
    >>> get_index('q0', ['q0', 'q1'])
    1
    >>> get_index('q1', ['q0', 'q1'])
    0
    """
    nq = len(labels)
    return flip_index(labels.index(target), nq)

def get_indices(targets, labels):
    return [get_index(t, labels) for t in targets]

def nonblank(args):
    for line in args:
        line = trim(line)
        if line.isspace():
            continue
        yield line
    return

def fullsplit(line):
    words = line.split()
    rest = ' '.join(words[1:])
    return fixcommand(words[0]), [s.strip() for s in rest.split(',')]

def fixcommand(c):
    """Fix Qasm command names.

    Remove all of forbidden characters from command c, and
    replace 'def' with 'qdef'.
    """
    forbidden_characters = ['-']
    c = c.lower()
    for char in forbidden_characters:
        c = c.replace(char, '')
    if c == 'def':
        return 'qdef'
    return c

def stripquotes(s):
    """Replace explicit quotes in a string.

    >>> from sympy.physics.quantum.qasm import stripquotes
    >>> stripquotes("'S'") == 'S'
    True
    >>> stripquotes('"S"') == 'S'
    True
    >>> stripquotes('S') == 'S'
    True
    """
    s = s.replace('"', '') # Remove second set of quotes?
    s = s.replace("'", '')
    return s

class Qasm:
    """Class to form objects from Qasm lines

    >>> from sympy.physics.quantum.qasm import Qasm
    >>> q = Qasm('qubit q0', 'qubit q1', 'h q0', 'cnot q0,q1')
    >>> q.get_circuit()
    CNOT(1,0)*H(1)
    >>> q = Qasm('qubit q0', 'qubit q1', 'cnot q0,q1', 'cnot q1,q0', 'cnot q0,q1')
    >>> q.get_circuit()
    CNOT(1,0)*CNOT(0,1)*CNOT(1,0)
    """
    def __init__(self, *args, **kwargs):
        self.defs = {}
        self.circuit = []
        self.labels = []
        self.inits = {}
        self.add(*args)
        self.kwargs = kwargs

    def add(self, *lines):
        for line in nonblank(lines):
            command, rest = fullsplit(line)
            if self.defs.get(command): #defs come first, since you can override built-in
                function = self.defs.get(command)
                indices = self.indices(rest)
                if len(indices) == 1:
                    self.circuit.append(function(indices[0]))
                else:
                    self.circuit.append(function(indices[:-1], indices[-1]))
            elif hasattr(self, command):
                function = getattr(self, command)
                function(*rest)
            else:
                print("Function %s not defined. Skipping" % command)

    def get_circuit(self):
        return prod(reversed(self.circuit))

    def get_labels(self):
        return list(reversed(self.labels))

    def plot(self):
        from sympy.physics.quantum.circuitplot import CircuitPlot
        circuit, labels = self.get_circuit(), self.get_labels()
        CircuitPlot(circuit, len(labels), labels=labels, inits=self.inits)

    def qubit(self, arg, init=None):
        self.labels.append(arg)
        if init: self.inits[arg] = init

    def indices(self, args):
        return get_indices(args, self.labels)

    def index(self, arg):
        return get_index(arg, self.labels)

    def nop(self, *args):
        pass

    def x(self, arg):
        self.circuit.append(X(self.index(arg)))

    def z(self, arg):
        self.circuit.append(Z(self.index(arg)))

    def h(self, arg):
        self.circuit.append(H(self.index(arg)))

    def s(self, arg):
        self.circuit.append(S(self.index(arg)))

    def t(self, arg):
        self.circuit.append(T(self.index(arg)))

    def measure(self, arg):
        self.circuit.append(Mz(self.index(arg)))

    def cnot(self, a1, a2):
        self.circuit.append(CNOT(*self.indices([a1, a2])))

    def swap(self, a1, a2):
        self.circuit.append(SWAP(*self.indices([a1, a2])))

    def cphase(self, a1, a2):
        self.circuit.append(CPHASE(*self.indices([a1, a2])))

    def toffoli(self, a1, a2, a3):
        i1, i2, i3 = self.indices([a1, a2, a3])
        self.circuit.append(CGateS((i1, i2), X(i3)))

    def cx(self, a1, a2):
        fi, fj = self.indices([a1, a2])
        self.circuit.append(CGate(fi, X(fj)))

    def cz(self, a1, a2):
        fi, fj = self.indices([a1, a2])
        self.circuit.append(CGate(fi, Z(fj)))

    def defbox(self, *args):
        print("defbox not supported yet. Skipping: ", args)

    def qdef(self, name, ncontrols, symbol):
        from sympy.physics.quantum.circuitplot import CreateOneQubitGate, CreateCGate
        ncontrols = int(ncontrols)
        command = fixcommand(name)
        symbol = stripquotes(symbol)
        if ncontrols > 0:
            self.defs[command] = CreateCGate(symbol)
        else:
            self.defs[command] = CreateOneQubitGate(symbol)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\printing\tensorflow.py ===
import sympy.codegen
import sympy.codegen.cfunctions
from sympy.external.importtools import version_tuple
from collections.abc import Iterable

from sympy.core.mul import Mul
from sympy.core.singleton import S
from sympy.codegen.cfunctions import Sqrt
from sympy.external import import_module
from sympy.printing.precedence import PRECEDENCE
from sympy.printing.pycode import AbstractPythonCodePrinter, ArrayPrinter
import sympy

tensorflow = import_module('tensorflow')

class TensorflowPrinter(ArrayPrinter, AbstractPythonCodePrinter):
    """
    Tensorflow printer which handles vectorized piecewise functions,
    logical operators, max/min, and relational operators.
    """
    printmethod = "_tensorflowcode"

    mapping = {
        sympy.Abs: "tensorflow.math.abs",
        sympy.sign: "tensorflow.math.sign",

        # XXX May raise error for ints.
        sympy.ceiling: "tensorflow.math.ceil",
        sympy.floor: "tensorflow.math.floor",
        sympy.log: "tensorflow.math.log",
        sympy.exp: "tensorflow.math.exp",
        Sqrt: "tensorflow.math.sqrt",
        sympy.cos: "tensorflow.math.cos",
        sympy.acos: "tensorflow.math.acos",
        sympy.sin: "tensorflow.math.sin",
        sympy.asin: "tensorflow.math.asin",
        sympy.tan: "tensorflow.math.tan",
        sympy.atan: "tensorflow.math.atan",
        sympy.atan2: "tensorflow.math.atan2",
        # XXX Also may give NaN for complex results.
        sympy.cosh: "tensorflow.math.cosh",
        sympy.acosh: "tensorflow.math.acosh",
        sympy.sinh: "tensorflow.math.sinh",
        sympy.asinh: "tensorflow.math.asinh",
        sympy.tanh: "tensorflow.math.tanh",
        sympy.atanh: "tensorflow.math.atanh",

        sympy.re: "tensorflow.math.real",
        sympy.im: "tensorflow.math.imag",
        sympy.arg: "tensorflow.math.angle",

        # XXX May raise error for ints and complexes
        sympy.erf: "tensorflow.math.erf",
        sympy.loggamma: "tensorflow.math.lgamma",

        sympy.Eq: "tensorflow.math.equal",
        sympy.Ne: "tensorflow.math.not_equal",
        sympy.StrictGreaterThan: "tensorflow.math.greater",
        sympy.StrictLessThan: "tensorflow.math.less",
        sympy.LessThan: "tensorflow.math.less_equal",
        sympy.GreaterThan: "tensorflow.math.greater_equal",

        sympy.And: "tensorflow.math.logical_and",
        sympy.Or: "tensorflow.math.logical_or",
        sympy.Not: "tensorflow.math.logical_not",
        sympy.Max: "tensorflow.math.maximum",
        sympy.Min: "tensorflow.math.minimum",

        # Matrices
        sympy.MatAdd: "tensorflow.math.add",
        sympy.HadamardProduct: "tensorflow.math.multiply",
        sympy.Trace: "tensorflow.linalg.trace",

        # XXX May raise error for integer matrices.
        sympy.Determinant : "tensorflow.linalg.det",
    }

    _default_settings = dict(
        AbstractPythonCodePrinter._default_settings,
        tensorflow_version=None
    )

    def __init__(self, settings=None):
        super().__init__(settings)

        version = self._settings['tensorflow_version']
        if version is None and tensorflow:
            version = tensorflow.__version__
        self.tensorflow_version = version

    def _print_Function(self, expr):
        op = self.mapping.get(type(expr), None)
        if op is None:
            return super()._print_Basic(expr)
        children = [self._print(arg) for arg in expr.args]
        if len(children) == 1:
            return "%s(%s)" % (
                self._module_format(op),
                children[0]
            )
        else:
            return self._expand_fold_binary_op(op, children)

    _print_Expr = _print_Function
    _print_Application = _print_Function
    _print_MatrixExpr = _print_Function
    # TODO: a better class structure would avoid this mess:
    _print_Relational = _print_Function
    _print_Not = _print_Function
    _print_And = _print_Function
    _print_Or = _print_Function
    _print_HadamardProduct = _print_Function
    _print_Trace = _print_Function
    _print_Determinant = _print_Function

    def _print_Inverse(self, expr):
        op = self._module_format('tensorflow.linalg.inv')
        return "{}({})".format(op, self._print(expr.arg))

    def _print_Transpose(self, expr):
        version = self.tensorflow_version
        if version and version_tuple(version) < version_tuple('1.14'):
            op = self._module_format('tensorflow.matrix_transpose')
        else:
            op = self._module_format('tensorflow.linalg.matrix_transpose')
        return "{}({})".format(op, self._print(expr.arg))

    def _print_Derivative(self, expr):
        variables = expr.variables
        if any(isinstance(i, Iterable) for i in variables):
            raise NotImplementedError("derivation by multiple variables is not supported")
        def unfold(expr, args):
            if not args:
                return self._print(expr)
            return "%s(%s, %s)[0]" % (
                    self._module_format("tensorflow.gradients"),
                    unfold(expr, args[:-1]),
                    self._print(args[-1]),
                )
        return unfold(expr.expr, variables)

    def _print_Piecewise(self, expr):
        version = self.tensorflow_version
        if version and version_tuple(version) < version_tuple('1.0'):
            tensorflow_piecewise = "tensorflow.select"
        else:
            tensorflow_piecewise = "tensorflow.where"

        from sympy.functions.elementary.piecewise import Piecewise
        e, cond = expr.args[0].args
        if len(expr.args) == 1:
            return '{}({}, {}, {})'.format(
                self._module_format(tensorflow_piecewise),
                self._print(cond),
                self._print(e),
                0)

        return '{}({}, {}, {})'.format(
            self._module_format(tensorflow_piecewise),
            self._print(cond),
            self._print(e),
            self._print(Piecewise(*expr.args[1:])))

    def _print_Pow(self, expr):
        # XXX May raise error for
        # int**float or int**complex or float**complex
        base, exp = expr.args
        if expr.exp == S.Half:
            return "{}({})".format(
                self._module_format("tensorflow.math.sqrt"), self._print(base))
        return "{}({}, {})".format(
            self._module_format("tensorflow.math.pow"),
            self._print(base), self._print(exp))

    def _print_MatrixBase(self, expr):
        tensorflow_f = "tensorflow.Variable" if expr.free_symbols else "tensorflow.constant"
        data = "["+", ".join(["["+", ".join([self._print(j) for j in i])+"]" for i in expr.tolist()])+"]"
        return "%s(%s)" % (
            self._module_format(tensorflow_f),
            data,
        )

    def _print_MatMul(self, expr):
        from sympy.matrices.expressions import MatrixExpr
        mat_args = [arg for arg in expr.args if isinstance(arg, MatrixExpr)]
        args = [arg for arg in expr.args if arg not in mat_args]
        if args:
            return "%s*%s" % (
                self.parenthesize(Mul.fromiter(args), PRECEDENCE["Mul"]),
                self._expand_fold_binary_op(
                    "tensorflow.linalg.matmul", mat_args)
            )
        else:
            return self._expand_fold_binary_op(
                "tensorflow.linalg.matmul", mat_args)

    def _print_MatPow(self, expr):
        return self._expand_fold_binary_op(
            "tensorflow.linalg.matmul", [expr.base]*expr.exp)

    def _print_CodeBlock(self, expr):
        # TODO: is this necessary?
        ret = []
        for subexpr in expr.args:
            ret.append(self._print(subexpr))
        return "\n".join(ret)

    def _print_isnan(self, exp):
        return f'tensorflow.math.is_nan({self._print(*exp.args)})'

    def _print_isinf(self, exp):
        return f'tensorflow.math.is_inf({self._print(*exp.args)})'

    _module = "tensorflow"
    _einsum = "linalg.einsum"
    _add = "math.add"
    _transpose = "transpose"
    _ones = "ones"
    _zeros = "zeros"


def tensorflow_code(expr, **settings):
    printer = TensorflowPrinter(settings)
    return printer.doprint(expr)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\tools\ort_format_model\ort_flatbuffers_py\fbs\Model.py ===
# automatically generated by the FlatBuffers compiler, do not modify

# namespace: fbs

import flatbuffers
from flatbuffers.compat import import_numpy
np = import_numpy()

class Model(object):
    __slots__ = ['_tab']

    @classmethod
    def GetRootAs(cls, buf, offset=0):
        n = flatbuffers.encode.Get(flatbuffers.packer.uoffset, buf, offset)
        x = Model()
        x.Init(buf, n + offset)
        return x

    @classmethod
    def GetRootAsModel(cls, buf, offset=0):
        """This method is deprecated. Please switch to GetRootAs."""
        return cls.GetRootAs(buf, offset)
    @classmethod
    def ModelBufferHasIdentifier(cls, buf, offset, size_prefixed=False):
        return flatbuffers.util.BufferHasIdentifier(buf, offset, b"\x4F\x52\x54\x4D", size_prefixed=size_prefixed)

    # Model
    def Init(self, buf, pos):
        self._tab = flatbuffers.table.Table(buf, pos)

    # Model
    def IrVersion(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(4))
        if o != 0:
            return self._tab.Get(flatbuffers.number_types.Int64Flags, o + self._tab.Pos)
        return 0

    # Model
    def OpsetImport(self, j):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        if o != 0:
            x = self._tab.Vector(o)
            x += flatbuffers.number_types.UOffsetTFlags.py_type(j) * 4
            x = self._tab.Indirect(x)
            from ort_flatbuffers_py.fbs.OperatorSetId import OperatorSetId
            obj = OperatorSetId()
            obj.Init(self._tab.Bytes, x)
            return obj
        return None

    # Model
    def OpsetImportLength(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        if o != 0:
            return self._tab.VectorLen(o)
        return 0

    # Model
    def OpsetImportIsNone(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(6))
        return o == 0

    # Model
    def ProducerName(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(8))
        if o != 0:
            return self._tab.String(o + self._tab.Pos)
        return None

    # Model
    def ProducerVersion(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(10))
        if o != 0:
            return self._tab.String(o + self._tab.Pos)
        return None

    # Model
    def Domain(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(12))
        if o != 0:
            return self._tab.String(o + self._tab.Pos)
        return None

    # Model
    def ModelVersion(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(14))
        if o != 0:
            return self._tab.Get(flatbuffers.number_types.Int64Flags, o + self._tab.Pos)
        return 0

    # Model
    def DocString(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(16))
        if o != 0:
            return self._tab.String(o + self._tab.Pos)
        return None

    # Model
    def Graph(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(18))
        if o != 0:
            x = self._tab.Indirect(o + self._tab.Pos)
            from ort_flatbuffers_py.fbs.Graph import Graph
            obj = Graph()
            obj.Init(self._tab.Bytes, x)
            return obj
        return None

    # Model
    def GraphDocString(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(20))
        if o != 0:
            return self._tab.String(o + self._tab.Pos)
        return None

    # Model
    def MetadataProps(self, j):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(22))
        if o != 0:
            x = self._tab.Vector(o)
            x += flatbuffers.number_types.UOffsetTFlags.py_type(j) * 4
            x = self._tab.Indirect(x)
            from ort_flatbuffers_py.fbs.StringStringEntry import StringStringEntry
            obj = StringStringEntry()
            obj.Init(self._tab.Bytes, x)
            return obj
        return None

    # Model
    def MetadataPropsLength(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(22))
        if o != 0:
            return self._tab.VectorLen(o)
        return 0

    # Model
    def MetadataPropsIsNone(self):
        o = flatbuffers.number_types.UOffsetTFlags.py_type(self._tab.Offset(22))
        return o == 0

def ModelStart(builder):
    builder.StartObject(10)

def Start(builder):
    ModelStart(builder)

def ModelAddIrVersion(builder, irVersion):
    builder.PrependInt64Slot(0, irVersion, 0)

def AddIrVersion(builder, irVersion):
    ModelAddIrVersion(builder, irVersion)

def ModelAddOpsetImport(builder, opsetImport):
    builder.PrependUOffsetTRelativeSlot(1, flatbuffers.number_types.UOffsetTFlags.py_type(opsetImport), 0)

def AddOpsetImport(builder, opsetImport):
    ModelAddOpsetImport(builder, opsetImport)

def ModelStartOpsetImportVector(builder, numElems):
    return builder.StartVector(4, numElems, 4)

def StartOpsetImportVector(builder, numElems: int) -> int:
    return ModelStartOpsetImportVector(builder, numElems)

def ModelAddProducerName(builder, producerName):
    builder.PrependUOffsetTRelativeSlot(2, flatbuffers.number_types.UOffsetTFlags.py_type(producerName), 0)

def AddProducerName(builder, producerName):
    ModelAddProducerName(builder, producerName)

def ModelAddProducerVersion(builder, producerVersion):
    builder.PrependUOffsetTRelativeSlot(3, flatbuffers.number_types.UOffsetTFlags.py_type(producerVersion), 0)

def AddProducerVersion(builder, producerVersion):
    ModelAddProducerVersion(builder, producerVersion)

def ModelAddDomain(builder, domain):
    builder.PrependUOffsetTRelativeSlot(4, flatbuffers.number_types.UOffsetTFlags.py_type(domain), 0)

def AddDomain(builder, domain):
    ModelAddDomain(builder, domain)

def ModelAddModelVersion(builder, modelVersion):
    builder.PrependInt64Slot(5, modelVersion, 0)

def AddModelVersion(builder, modelVersion):
    ModelAddModelVersion(builder, modelVersion)

def ModelAddDocString(builder, docString):
    builder.PrependUOffsetTRelativeSlot(6, flatbuffers.number_types.UOffsetTFlags.py_type(docString), 0)

def AddDocString(builder, docString):
    ModelAddDocString(builder, docString)

def ModelAddGraph(builder, graph):
    builder.PrependUOffsetTRelativeSlot(7, flatbuffers.number_types.UOffsetTFlags.py_type(graph), 0)

def AddGraph(builder, graph):
    ModelAddGraph(builder, graph)

def ModelAddGraphDocString(builder, graphDocString):
    builder.PrependUOffsetTRelativeSlot(8, flatbuffers.number_types.UOffsetTFlags.py_type(graphDocString), 0)

def AddGraphDocString(builder, graphDocString):
    ModelAddGraphDocString(builder, graphDocString)

def ModelAddMetadataProps(builder, metadataProps):
    builder.PrependUOffsetTRelativeSlot(9, flatbuffers.number_types.UOffsetTFlags.py_type(metadataProps), 0)

def AddMetadataProps(builder, metadataProps):
    ModelAddMetadataProps(builder, metadataProps)

def ModelStartMetadataPropsVector(builder, numElems):
    return builder.StartVector(4, numElems, 4)

def StartMetadataPropsVector(builder, numElems: int) -> int:
    return ModelStartMetadataPropsVector(builder, numElems)

def ModelEnd(builder):
    return builder.EndObject()

def End(builder):
    return ModelEnd(builder)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\rich\progress_bar.py ===
import math
from functools import lru_cache
from time import monotonic
from typing import Iterable, List, Optional

from .color import Color, blend_rgb
from .color_triplet import ColorTriplet
from .console import Console, ConsoleOptions, RenderResult
from .jupyter import JupyterMixin
from .measure import Measurement
from .segment import Segment
from .style import Style, StyleType

# Number of characters before 'pulse' animation repeats
PULSE_SIZE = 20


class ProgressBar(JupyterMixin):
    """Renders a (progress) bar. Used by rich.progress.

    Args:
        total (float, optional): Number of steps in the bar. Defaults to 100. Set to None to render a pulsing animation.
        completed (float, optional): Number of steps completed. Defaults to 0.
        width (int, optional): Width of the bar, or ``None`` for maximum width. Defaults to None.
        pulse (bool, optional): Enable pulse effect. Defaults to False. Will pulse if a None total was passed.
        style (StyleType, optional): Style for the bar background. Defaults to "bar.back".
        complete_style (StyleType, optional): Style for the completed bar. Defaults to "bar.complete".
        finished_style (StyleType, optional): Style for a finished bar. Defaults to "bar.finished".
        pulse_style (StyleType, optional): Style for pulsing bars. Defaults to "bar.pulse".
        animation_time (Optional[float], optional): Time in seconds to use for animation, or None to use system time.
    """

    def __init__(
        self,
        total: Optional[float] = 100.0,
        completed: float = 0,
        width: Optional[int] = None,
        pulse: bool = False,
        style: StyleType = "bar.back",
        complete_style: StyleType = "bar.complete",
        finished_style: StyleType = "bar.finished",
        pulse_style: StyleType = "bar.pulse",
        animation_time: Optional[float] = None,
    ):
        self.total = total
        self.completed = completed
        self.width = width
        self.pulse = pulse
        self.style = style
        self.complete_style = complete_style
        self.finished_style = finished_style
        self.pulse_style = pulse_style
        self.animation_time = animation_time

        self._pulse_segments: Optional[List[Segment]] = None

    def __repr__(self) -> str:
        return f"<Bar {self.completed!r} of {self.total!r}>"

    @property
    def percentage_completed(self) -> Optional[float]:
        """Calculate percentage complete."""
        if self.total is None:
            return None
        completed = (self.completed / self.total) * 100.0
        completed = min(100, max(0.0, completed))
        return completed

    @lru_cache(maxsize=16)
    def _get_pulse_segments(
        self,
        fore_style: Style,
        back_style: Style,
        color_system: str,
        no_color: bool,
        ascii: bool = False,
    ) -> List[Segment]:
        """Get a list of segments to render a pulse animation.

        Returns:
            List[Segment]: A list of segments, one segment per character.
        """
        bar = "-" if ascii else "━"
        segments: List[Segment] = []
        if color_system not in ("standard", "eight_bit", "truecolor") or no_color:
            segments += [Segment(bar, fore_style)] * (PULSE_SIZE // 2)
            segments += [Segment(" " if no_color else bar, back_style)] * (
                PULSE_SIZE - (PULSE_SIZE // 2)
            )
            return segments

        append = segments.append
        fore_color = (
            fore_style.color.get_truecolor()
            if fore_style.color
            else ColorTriplet(255, 0, 255)
        )
        back_color = (
            back_style.color.get_truecolor()
            if back_style.color
            else ColorTriplet(0, 0, 0)
        )
        cos = math.cos
        pi = math.pi
        _Segment = Segment
        _Style = Style
        from_triplet = Color.from_triplet

        for index in range(PULSE_SIZE):
            position = index / PULSE_SIZE
            fade = 0.5 + cos(position * pi * 2) / 2.0
            color = blend_rgb(fore_color, back_color, cross_fade=fade)
            append(_Segment(bar, _Style(color=from_triplet(color))))
        return segments

    def update(self, completed: float, total: Optional[float] = None) -> None:
        """Update progress with new values.

        Args:
            completed (float): Number of steps completed.
            total (float, optional): Total number of steps, or ``None`` to not change. Defaults to None.
        """
        self.completed = completed
        self.total = total if total is not None else self.total

    def _render_pulse(
        self, console: Console, width: int, ascii: bool = False
    ) -> Iterable[Segment]:
        """Renders the pulse animation.

        Args:
            console (Console): Console instance.
            width (int): Width in characters of pulse animation.

        Returns:
            RenderResult: [description]

        Yields:
            Iterator[Segment]: Segments to render pulse
        """
        fore_style = console.get_style(self.pulse_style, default="white")
        back_style = console.get_style(self.style, default="black")

        pulse_segments = self._get_pulse_segments(
            fore_style, back_style, console.color_system, console.no_color, ascii=ascii
        )
        segment_count = len(pulse_segments)
        current_time = (
            monotonic() if self.animation_time is None else self.animation_time
        )
        segments = pulse_segments * (int(width / segment_count) + 2)
        offset = int(-current_time * 15) % segment_count
        segments = segments[offset : offset + width]
        yield from segments

    def __rich_console__(
        self, console: Console, options: ConsoleOptions
    ) -> RenderResult:
        width = min(self.width or options.max_width, options.max_width)
        ascii = options.legacy_windows or options.ascii_only
        should_pulse = self.pulse or self.total is None
        if should_pulse:
            yield from self._render_pulse(console, width, ascii=ascii)
            return

        completed: Optional[float] = (
            min(self.total, max(0, self.completed)) if self.total is not None else None
        )

        bar = "-" if ascii else "━"
        half_bar_right = " " if ascii else "╸"
        half_bar_left = " " if ascii else "╺"
        complete_halves = (
            int(width * 2 * completed / self.total)
            if self.total and completed is not None
            else width * 2
        )
        bar_count = complete_halves // 2
        half_bar_count = complete_halves % 2
        style = console.get_style(self.style)
        is_finished = self.total is None or self.completed >= self.total
        complete_style = console.get_style(
            self.finished_style if is_finished else self.complete_style
        )
        _Segment = Segment
        if bar_count:
            yield _Segment(bar * bar_count, complete_style)
        if half_bar_count:
            yield _Segment(half_bar_right * half_bar_count, complete_style)

        if not console.no_color:
            remaining_bars = width - bar_count - half_bar_count
            if remaining_bars and console.color_system is not None:
                if not half_bar_count and bar_count:
                    yield _Segment(half_bar_left, style)
                    remaining_bars -= 1
                if remaining_bars:
                    yield _Segment(bar * remaining_bars, style)

    def __rich_measure__(
        self, console: Console, options: ConsoleOptions
    ) -> Measurement:
        return (
            Measurement(self.width, self.width)
            if self.width is not None
            else Measurement(4, options.max_width)
        )


if __name__ == "__main__":  # pragma: no cover
    console = Console()
    bar = ProgressBar(width=50, total=100)

    import time

    console.show_cursor(False)
    for n in range(0, 101, 1):
        bar.update(n)
        console.print(bar)
        console.file.write("\r")
        time.sleep(0.05)
    console.show_cursor(True)
    console.print()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\ntheory\egyptian_fraction.py ===
from sympy.core.containers import Tuple
from sympy.core.numbers import (Integer, Rational)
from sympy.core.singleton import S
import sympy.polys

from math import gcd


def egyptian_fraction(r, algorithm="Greedy"):
    """
    Return the list of denominators of an Egyptian fraction
    expansion [1]_ of the said rational `r`.

    Parameters
    ==========

    r : Rational or (p, q)
        a positive rational number, ``p/q``.
    algorithm : { "Greedy", "Graham Jewett", "Takenouchi", "Golomb" }, optional
        Denotes the algorithm to be used (the default is "Greedy").

    Examples
    ========

    >>> from sympy import Rational
    >>> from sympy.ntheory.egyptian_fraction import egyptian_fraction
    >>> egyptian_fraction(Rational(3, 7))
    [3, 11, 231]
    >>> egyptian_fraction((3, 7), "Graham Jewett")
    [7, 8, 9, 56, 57, 72, 3192]
    >>> egyptian_fraction((3, 7), "Takenouchi")
    [4, 7, 28]
    >>> egyptian_fraction((3, 7), "Golomb")
    [3, 15, 35]
    >>> egyptian_fraction((11, 5), "Golomb")
    [1, 2, 3, 4, 9, 234, 1118, 2580]

    See Also
    ========

    sympy.core.numbers.Rational

    Notes
    =====

    Currently the following algorithms are supported:

    1) Greedy Algorithm

       Also called the Fibonacci-Sylvester algorithm [2]_.
       At each step, extract the largest unit fraction less
       than the target and replace the target with the remainder.

       It has some distinct properties:

       a) Given `p/q` in lowest terms, generates an expansion of maximum
          length `p`. Even as the numerators get large, the number of
          terms is seldom more than a handful.

       b) Uses minimal memory.

       c) The terms can blow up (standard examples of this are 5/121 and
          31/311).  The denominator is at most squared at each step
          (doubly-exponential growth) and typically exhibits
          singly-exponential growth.

    2) Graham Jewett Algorithm

       The algorithm suggested by the result of Graham and Jewett.
       Note that this has a tendency to blow up: the length of the
       resulting expansion is always ``2**(x/gcd(x, y)) - 1``.  See [3]_.

    3) Takenouchi Algorithm

       The algorithm suggested by Takenouchi (1921).
       Differs from the Graham-Jewett algorithm only in the handling
       of duplicates.  See [3]_.

    4) Golomb's Algorithm

       A method given by Golumb (1962), using modular arithmetic and
       inverses.  It yields the same results as a method using continued
       fractions proposed by Bleicher (1972).  See [4]_.

    If the given rational is greater than or equal to 1, a greedy algorithm
    of summing the harmonic sequence 1/1 + 1/2 + 1/3 + ... is used, taking
    all the unit fractions of this sequence until adding one more would be
    greater than the given number.  This list of denominators is prefixed
    to the result from the requested algorithm used on the remainder.  For
    example, if r is 8/3, using the Greedy algorithm, we get [1, 2, 3, 4,
    5, 6, 7, 14, 420], where the beginning of the sequence, [1, 2, 3, 4, 5,
    6, 7] is part of the harmonic sequence summing to 363/140, leaving a
    remainder of 31/420, which yields [14, 420] by the Greedy algorithm.
    The result of egyptian_fraction(Rational(8, 3), "Golomb") is [1, 2, 3,
    4, 5, 6, 7, 14, 574, 2788, 6460, 11590, 33062, 113820], and so on.

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Egyptian_fraction
    .. [2] https://en.wikipedia.org/wiki/Greedy_algorithm_for_Egyptian_fractions
    .. [3] https://www.ics.uci.edu/~eppstein/numth/egypt/conflict.html
    .. [4] https://web.archive.org/web/20180413004012/https://ami.ektf.hu/uploads/papers/finalpdf/AMI_42_from129to134.pdf

    """

    if not isinstance(r, Rational):
        if isinstance(r, (Tuple, tuple)) and len(r) == 2:
            r = Rational(*r)
        else:
            raise ValueError("Value must be a Rational or tuple of ints")
    if r <= 0:
        raise ValueError("Value must be positive")

    # common cases that all methods agree on
    x, y = r.as_numer_denom()
    if y == 1 and x == 2:
        return [Integer(i) for i in [1, 2, 3, 6]]
    if x == y + 1:
        return [S.One, y]

    prefix, rem = egypt_harmonic(r)
    if rem == 0:
        return prefix
    # work in Python ints
    x, y = rem.p, rem.q
    # assert x < y and gcd(x, y) = 1

    if algorithm == "Greedy":
        postfix = egypt_greedy(x, y)
    elif algorithm == "Graham Jewett":
        postfix = egypt_graham_jewett(x, y)
    elif algorithm == "Takenouchi":
        postfix = egypt_takenouchi(x, y)
    elif algorithm == "Golomb":
        postfix = egypt_golomb(x, y)
    else:
        raise ValueError("Entered invalid algorithm")
    return prefix + [Integer(i) for i in postfix]


def egypt_greedy(x, y):
    # assumes gcd(x, y) == 1
    if x == 1:
        return [y]
    else:
        a = (-y) % x
        b = y*(y//x + 1)
        c = gcd(a, b)
        if c > 1:
            num, denom = a//c, b//c
        else:
            num, denom = a, b
        return [y//x + 1] + egypt_greedy(num, denom)


def egypt_graham_jewett(x, y):
    # assumes gcd(x, y) == 1
    l = [y] * x

    # l is now a list of integers whose reciprocals sum to x/y.
    # we shall now proceed to manipulate the elements of l without
    # changing the reciprocated sum until all elements are unique.

    while len(l) != len(set(l)):
        l.sort()  # so the list has duplicates. find a smallest pair
        for i in range(len(l) - 1):
            if l[i] == l[i + 1]:
                break
        # we have now identified a pair of identical
        # elements: l[i] and l[i + 1].
        # now comes the application of the result of graham and jewett:
        l[i + 1] = l[i] + 1
        # and we just iterate that until the list has no duplicates.
        l.append(l[i]*(l[i] + 1))
    return sorted(l)


def egypt_takenouchi(x, y):
    # assumes gcd(x, y) == 1
    # special cases for 3/y
    if x == 3:
        if y % 2 == 0:
            return [y//2, y]
        i = (y - 1)//2
        j = i + 1
        k = j + i
        return [j, k, j*k]
    l = [y] * x
    while len(l) != len(set(l)):
        l.sort()
        for i in range(len(l) - 1):
            if l[i] == l[i + 1]:
                break
        k = l[i]
        if k % 2 == 0:
            l[i] = l[i] // 2
            del l[i + 1]
        else:
            l[i], l[i + 1] = (k + 1)//2, k*(k + 1)//2
    return sorted(l)


def egypt_golomb(x, y):
    # assumes x < y and gcd(x, y) == 1
    if x == 1:
        return [y]
    xp = sympy.polys.ZZ.invert(int(x), int(y))
    rv = [xp*y]
    rv.extend(egypt_golomb((x*xp - 1)//y, xp))
    return sorted(rv)


def egypt_harmonic(r):
    # assumes r is Rational
    rv = []
    d = S.One
    acc = S.Zero
    while acc + 1/d <= r:
        acc += 1/d
        rv.append(d)
        d += 1
    return (rv, r - acc)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\stats\compound_rv.py ===
from sympy.concrete.summations import Sum
from sympy.core.basic import Basic
from sympy.core.function import Lambda
from sympy.core.symbol import Dummy
from sympy.integrals.integrals import Integral
from sympy.stats.rv import (NamedArgsMixin, random_symbols, _symbol_converter,
                        PSpace, RandomSymbol, is_random, Distribution)
from sympy.stats.crv import ContinuousDistribution, SingleContinuousPSpace
from sympy.stats.drv import DiscreteDistribution, SingleDiscretePSpace
from sympy.stats.frv import SingleFiniteDistribution, SingleFinitePSpace
from sympy.stats.crv_types import ContinuousDistributionHandmade
from sympy.stats.drv_types import DiscreteDistributionHandmade
from sympy.stats.frv_types import FiniteDistributionHandmade


class CompoundPSpace(PSpace):
    """
    A temporary Probability Space for the Compound Distribution. After
    Marginalization, this returns the corresponding Probability Space of the
    parent distribution.
    """

    def __new__(cls, s, distribution):
        s = _symbol_converter(s)
        if isinstance(distribution, ContinuousDistribution):
            return SingleContinuousPSpace(s, distribution)
        if isinstance(distribution, DiscreteDistribution):
            return SingleDiscretePSpace(s, distribution)
        if isinstance(distribution, SingleFiniteDistribution):
            return SingleFinitePSpace(s, distribution)
        if not isinstance(distribution, CompoundDistribution):
            raise ValueError("%s should be an isinstance of "
                        "CompoundDistribution"%(distribution))
        return Basic.__new__(cls, s, distribution)

    @property
    def value(self):
        return RandomSymbol(self.symbol, self)

    @property
    def symbol(self):
        return self.args[0]

    @property
    def is_Continuous(self):
        return self.distribution.is_Continuous

    @property
    def is_Finite(self):
        return self.distribution.is_Finite

    @property
    def is_Discrete(self):
        return self.distribution.is_Discrete

    @property
    def distribution(self):
        return self.args[1]

    @property
    def pdf(self):
        return self.distribution.pdf(self.symbol)

    @property
    def set(self):
        return self.distribution.set

    @property
    def domain(self):
        return self._get_newpspace().domain

    def _get_newpspace(self, evaluate=False):
        x = Dummy('x')
        parent_dist = self.distribution.args[0]
        func = Lambda(x, self.distribution.pdf(x, evaluate))
        new_pspace = self._transform_pspace(self.symbol, parent_dist, func)
        if new_pspace is not None:
            return new_pspace
        message = ("Compound Distribution for %s is not implemented yet" % str(parent_dist))
        raise NotImplementedError(message)

    def _transform_pspace(self, sym, dist, pdf):
        """
        This function returns the new pspace of the distribution using handmade
        Distributions and their corresponding pspace.
        """
        pdf = Lambda(sym, pdf(sym))
        _set = dist.set
        if isinstance(dist, ContinuousDistribution):
            return SingleContinuousPSpace(sym, ContinuousDistributionHandmade(pdf, _set))
        elif isinstance(dist, DiscreteDistribution):
            return SingleDiscretePSpace(sym, DiscreteDistributionHandmade(pdf, _set))
        elif isinstance(dist, SingleFiniteDistribution):
            dens = {k: pdf(k) for k in _set}
            return SingleFinitePSpace(sym, FiniteDistributionHandmade(dens))

    def compute_density(self, expr, *, compound_evaluate=True, **kwargs):
        new_pspace = self._get_newpspace(compound_evaluate)
        expr = expr.subs({self.value: new_pspace.value})
        return new_pspace.compute_density(expr, **kwargs)

    def compute_cdf(self, expr, *, compound_evaluate=True, **kwargs):
        new_pspace = self._get_newpspace(compound_evaluate)
        expr = expr.subs({self.value: new_pspace.value})
        return new_pspace.compute_cdf(expr, **kwargs)

    def compute_expectation(self, expr, rvs=None, evaluate=False, **kwargs):
        new_pspace = self._get_newpspace(evaluate)
        expr = expr.subs({self.value: new_pspace.value})
        if rvs:
            rvs = rvs.subs({self.value: new_pspace.value})
        if isinstance(new_pspace, SingleFinitePSpace):
            return new_pspace.compute_expectation(expr, rvs, **kwargs)
        return new_pspace.compute_expectation(expr, rvs, evaluate, **kwargs)

    def probability(self, condition, *, compound_evaluate=True, **kwargs):
        new_pspace = self._get_newpspace(compound_evaluate)
        condition = condition.subs({self.value: new_pspace.value})
        return new_pspace.probability(condition)

    def conditional_space(self, condition, *, compound_evaluate=True, **kwargs):
        new_pspace = self._get_newpspace(compound_evaluate)
        condition = condition.subs({self.value: new_pspace.value})
        return new_pspace.conditional_space(condition)


class CompoundDistribution(Distribution, NamedArgsMixin):
    """
    Class for Compound Distributions.

    Parameters
    ==========

    dist : Distribution
        Distribution must contain a random parameter

    Examples
    ========

    >>> from sympy.stats.compound_rv import CompoundDistribution
    >>> from sympy.stats.crv_types import NormalDistribution
    >>> from sympy.stats import Normal
    >>> from sympy.abc import x
    >>> X = Normal('X', 2, 4)
    >>> N = NormalDistribution(X, 4)
    >>> C = CompoundDistribution(N)
    >>> C.set
    Interval(-oo, oo)
    >>> C.pdf(x, evaluate=True).simplify()
    exp(-x**2/64 + x/16 - 1/16)/(8*sqrt(pi))

    References
    ==========

    .. [1] https://en.wikipedia.org/wiki/Compound_probability_distribution

    """

    def __new__(cls, dist):
        if not isinstance(dist, (ContinuousDistribution,
                SingleFiniteDistribution, DiscreteDistribution)):
            message = "Compound Distribution for %s is not implemented yet" % str(dist)
            raise NotImplementedError(message)
        if not cls._compound_check(dist):
            return dist
        return Basic.__new__(cls, dist)

    @property
    def set(self):
        return self.args[0].set

    @property
    def is_Continuous(self):
        return isinstance(self.args[0], ContinuousDistribution)

    @property
    def is_Finite(self):
        return isinstance(self.args[0], SingleFiniteDistribution)

    @property
    def is_Discrete(self):
        return isinstance(self.args[0], DiscreteDistribution)

    def pdf(self, x, evaluate=False):
        dist = self.args[0]
        randoms = [rv for rv in dist.args if is_random(rv)]
        if isinstance(dist, SingleFiniteDistribution):
            y = Dummy('y', integer=True, negative=False)
            expr = dist.pmf(y)
        else:
            y = Dummy('y')
            expr = dist.pdf(y)
        for rv in randoms:
            expr = self._marginalise(expr, rv, evaluate)
        return Lambda(y, expr)(x)

    def _marginalise(self, expr, rv, evaluate):
        if isinstance(rv.pspace.distribution, SingleFiniteDistribution):
            rv_dens = rv.pspace.distribution.pmf(rv)
        else:
            rv_dens = rv.pspace.distribution.pdf(rv)
        rv_dom = rv.pspace.domain.set
        if rv.pspace.is_Discrete or rv.pspace.is_Finite:
            expr = Sum(expr*rv_dens, (rv, rv_dom._inf,
                    rv_dom._sup))
        else:
            expr = Integral(expr*rv_dens, (rv, rv_dom._inf,
                    rv_dom._sup))
        if evaluate:
            return expr.doit()
        return expr

    @classmethod
    def _compound_check(self, dist):
        """
        Checks if the given distribution contains random parameters.
        """
        randoms = []
        for arg in dist.args:
            randoms.extend(random_symbols(arg))
        if len(randoms) == 0:
            return False
        return True

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\trio\_core\_entry_queue.py ===
from __future__ import annotations

import threading
from collections import deque
from collections.abc import Callable
from typing import TYPE_CHECKING, NoReturn

import attrs

from .. import _core
from .._util import NoPublicConstructor, final
from ._wakeup_socketpair import WakeupSocketpair

if TYPE_CHECKING:
    from typing_extensions import TypeVarTuple, Unpack

    PosArgsT = TypeVarTuple("PosArgsT")

Function = Callable[..., object]  # type: ignore[explicit-any]
Job = tuple[Function, tuple[object, ...]]


@attrs.define
class EntryQueue:
    # This used to use a queue.Queue. but that was broken, because Queues are
    # implemented in Python, and not reentrant -- so it was thread-safe, but
    # not signal-safe. deque is implemented in C, so each operation is atomic
    # WRT threads (and this is guaranteed in the docs), AND each operation is
    # atomic WRT signal delivery (signal handlers can run on either side, but
    # not *during* a deque operation). dict makes similar guarantees - and
    # it's even ordered!
    queue: deque[Job] = attrs.Factory(deque)
    idempotent_queue: dict[Job, None] = attrs.Factory(dict)

    wakeup: WakeupSocketpair = attrs.Factory(WakeupSocketpair)
    done: bool = False
    # Must be a reentrant lock, because it's acquired from signal handlers.
    # RLock is signal-safe as of cpython 3.2. NB that this does mean that the
    # lock is effectively *disabled* when we enter from signal context. The
    # way we use the lock this is OK though, because when
    # run_sync_soon is called from a signal it's atomic WRT the
    # main thread -- it just might happen at some inconvenient place. But if
    # you look at the one place where the main thread holds the lock, it's
    # just to make 1 assignment, so that's atomic WRT a signal anyway.
    lock: threading.RLock = attrs.Factory(threading.RLock)

    async def task(self) -> None:
        assert _core.currently_ki_protected()
        # RLock has two implementations: a signal-safe version in _thread, and
        # and signal-UNsafe version in threading. We need the signal safe
        # version. Python 3.2 and later should always use this anyway, but,
        # since the symptoms if this goes wrong are just "weird rare
        # deadlocks", then let's make a little check.
        # See:
        #     https://bugs.python.org/issue13697#msg237140
        assert self.lock.__class__.__module__ == "_thread"

        def run_cb(job: Job) -> None:
            # We run this with KI protection enabled; it's the callback's
            # job to disable it if it wants it disabled. Exceptions are
            # treated like system task exceptions (i.e., converted into
            # TrioInternalError and cause everything to shut down).
            sync_fn, args = job
            try:
                sync_fn(*args)
            except BaseException as exc:

                async def kill_everything(  # noqa: RUF029  # await not used
                    exc: BaseException,
                ) -> NoReturn:
                    raise exc

                try:
                    _core.spawn_system_task(kill_everything, exc)
                except RuntimeError:
                    # We're quite late in the shutdown process and the
                    # system nursery is already closed.
                    # TODO(2020-06): this is a gross hack and should
                    # be fixed soon when we address #1607.
                    parent_nursery = _core.current_task().parent_nursery
                    if parent_nursery is None:
                        raise AssertionError(
                            "Internal error: `parent_nursery` should never be `None`",
                        ) from exc  # pragma: no cover
                    parent_nursery.start_soon(kill_everything, exc)

        # This has to be carefully written to be safe in the face of new items
        # being queued while we iterate, and to do a bounded amount of work on
        # each pass:
        def run_all_bounded() -> None:
            for _ in range(len(self.queue)):
                run_cb(self.queue.popleft())
            for job in list(self.idempotent_queue):
                del self.idempotent_queue[job]
                run_cb(job)

        try:
            while True:
                run_all_bounded()
                if not self.queue and not self.idempotent_queue:
                    await self.wakeup.wait_woken()
                else:
                    await _core.checkpoint()
        except _core.Cancelled:
            # Keep the work done with this lock held as minimal as possible,
            # because it doesn't protect us against concurrent signal delivery
            # (see the comment above). Notice that this code would still be
            # correct if written like:
            #   self.done = True
            #   with self.lock:
            #       pass
            # because all we want is to force run_sync_soon
            # to either be completely before or completely after the write to
            # done. That's why we don't need the lock to protect
            # against signal handlers.
            with self.lock:
                self.done = True
            # No more jobs will be submitted, so just clear out any residual
            # ones:
            run_all_bounded()
            assert not self.queue
            assert not self.idempotent_queue

    def close(self) -> None:
        self.wakeup.close()

    def size(self) -> int:
        return len(self.queue) + len(self.idempotent_queue)

    def run_sync_soon(
        self,
        sync_fn: Callable[[Unpack[PosArgsT]], object],
        *args: Unpack[PosArgsT],
        idempotent: bool = False,
    ) -> None:
        with self.lock:
            if self.done:
                raise _core.RunFinishedError("run() has exited")
            # We have to hold the lock all the way through here, because
            # otherwise the main thread might exit *while* we're doing these
            # calls, and then our queue item might not be processed, or the
            # wakeup call might trigger an OSError b/c the IO manager has
            # already been shut down.
            if idempotent:
                self.idempotent_queue[sync_fn, args] = None
            else:
                self.queue.append((sync_fn, args))
            self.wakeup.wakeup_thread_and_signal_safe()


@final
@attrs.define(eq=False)
class TrioToken(metaclass=NoPublicConstructor):
    """An opaque object representing a single call to :func:`trio.run`.

    It has no public constructor; instead, see :func:`current_trio_token`.

    This object has two uses:

    1. It lets you re-enter the Trio run loop from external threads or signal
       handlers. This is the low-level primitive that :func:`trio.to_thread`
       and `trio.from_thread` use to communicate with worker threads, that
       `trio.open_signal_receiver` uses to receive notifications about
       signals, and so forth.

    2. Each call to :func:`trio.run` has exactly one associated
       :class:`TrioToken` object, so you can use it to identify a particular
       call.

    """

    _reentry_queue: EntryQueue

    def run_sync_soon(
        self,
        sync_fn: Callable[[Unpack[PosArgsT]], object],
        *args: Unpack[PosArgsT],
        idempotent: bool = False,
    ) -> None:
        """Schedule a call to ``sync_fn(*args)`` to occur in the context of a
        Trio task.

        This is safe to call from the main thread, from other threads, and
        from signal handlers. This is the fundamental primitive used to
        re-enter the Trio run loop from outside of it.

        The call will happen "soon", but there's no guarantee about exactly
        when, and no mechanism provided for finding out when it's happened.
        If you need this, you'll have to build your own.

        The call is effectively run as part of a system task (see
        :func:`~trio.lowlevel.spawn_system_task`). In particular this means
        that:

        * :exc:`KeyboardInterrupt` protection is *enabled* by default; if
          you want ``sync_fn`` to be interruptible by control-C, then you
          need to use :func:`~trio.lowlevel.disable_ki_protection`
          explicitly.

        * If ``sync_fn`` raises an exception, then it's converted into a
          :exc:`~trio.TrioInternalError` and *all* tasks are cancelled. You
          should be careful that ``sync_fn`` doesn't crash.

        All calls with ``idempotent=False`` are processed in strict
        first-in first-out order.

        If ``idempotent=True``, then ``sync_fn`` and ``args`` must be
        hashable, and Trio will make a best-effort attempt to discard any
        call submission which is equal to an already-pending call. Trio
        will process these in first-in first-out order.

        Any ordering guarantees apply separately to ``idempotent=False``
        and ``idempotent=True`` calls; there's no rule for how calls in the
        different categories are ordered with respect to each other.

        :raises trio.RunFinishedError:
              if the associated call to :func:`trio.run`
              has already exited. (Any call that *doesn't* raise this error
              is guaranteed to be fully processed before :func:`trio.run`
              exits.)

        """
        self._reentry_queue.run_sync_soon(sync_fn, *args, idempotent=idempotent)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\onnxruntime\transformers\fusion_conformer_attention.py ===
# -------------------------------------------------------------------------
# Copyright (c) Microsoft Corporation.  All rights reserved.
# Licensed under the MIT License.
# --------------------------------------------------------------------------
import logging

from fusion_attention import AttentionMask, FusionAttention
from onnx_model import OnnxModel

logger = logging.getLogger(__name__)


class FusionConformerAttention(FusionAttention):
    """
    Fuse Conformer Attention subgraph into one MultiHeadAttention node.
    """

    def __init__(
        self,
        model: OnnxModel,
        hidden_size: int,
        num_heads: int,
        attention_mask: AttentionMask,
    ):
        super().__init__(model, hidden_size, num_heads, attention_mask)

    def fuse(self, normalize_node, input_name_to_nodes, output_name_to_node):
        # SkipLayerNormalization has two inputs, and one of them is the root input for attention.
        qkv_nodes = self.model.match_parent_path(
            normalize_node,
            ["Add", "MatMul", "Reshape", "Transpose", "MatMul"],
            [1, None, 0, 0, 0],
        )
        if qkv_nodes is None:
            logger.debug("fuse_conformer_attention: failed to match qkv path")
            return

        reshape_qkv, transpose_qkv, matmul_qkv = qkv_nodes[-3], qkv_nodes[-2], qkv_nodes[-1]

        past_v, present_v = "", ""
        v_nodes = self.model.match_parent_path(
            matmul_qkv,
            ["Concat", "Transpose", "Reshape", "Add", "MatMul"],
            [1, 1, 0, 0, 1],
        )
        if v_nodes is None:
            v_nodes = self.model.match_parent_path(
                matmul_qkv,
                ["Transpose", "Reshape", "Add", "MatMul"],
                [1, 0, 0, 0],
            )
            if v_nodes is None:
                logger.debug("fuse_conformer_attention: failed to match v path")
                return
        else:
            concat_v = v_nodes[0]
            concat_parent = self.model.get_parent(concat_v, 0, None)
            present_v = concat_v.output[0]
            past_v = concat_parent.output[0]

        add_v, matmul_v = v_nodes[-2], v_nodes[-1]

        attn_mask = ""
        qk_nodes = self.model.match_parent_path(
            matmul_qkv,
            ["Softmax", "Add", "MatMul"],
            [0, 0, 0],
        )
        if qk_nodes is None:
            qk_nodes = self.model.match_parent_path(
                matmul_qkv,
                ["Where", "Softmax", "Where", "Add", "MatMul"],
                [0, 2, 0, 2, 0],
            )
            if qk_nodes is None:
                logger.debug("fuse_conformer_attention: failed to match qk path")
                return

            where_qk = qk_nodes[2]
            mask_nodes = self.model.match_parent_path(
                where_qk,
                ["Equal", "Unsqueeze", "Cast"],
                [0, 0, 0],
            )
            if mask_nodes is not None:
                attn_mask = mask_nodes[-1].output[0]

        add_qk, matmul_qk = qk_nodes[-2], qk_nodes[-1]

        q_nodes = self.model.match_parent_path(
            matmul_qk,
            ["Div", "Transpose", "Reshape", "Add", "MatMul"],
            [0, 0, 0, 0, 1],
        )
        if q_nodes is None:
            q_nodes = self.model.match_parent_path(
                matmul_qk,
                ["Mul", "Transpose", "Reshape", "Add", "MatMul"],
                [0, 0, 0, 0, 0],
            )
            if q_nodes is None:
                logger.debug("fuse_conformer_attention: failed to match q path")
                return

        reshape_q, add_q, matmul_q = q_nodes[-3], q_nodes[-2], q_nodes[-1]

        extra_q_nodes = self.model.match_parent_path(
            add_qk,
            ["Reshape", "Transpose", "MatMul", "Transpose", "Reshape", "Div"],
            [1, 0, 0, 0, 0, 0],
        )
        if extra_q_nodes is not None and q_nodes[0] != extra_q_nodes[-1]:
            logger.debug("fuse_conformer_attention: failed to match extra q path")
            return

        past_k, present_k = "", ""
        k_nodes = self.model.match_parent_path(
            matmul_qk,
            ["Transpose", "Concat", "Transpose", "Reshape", "Add", "MatMul"],
            [1, 0, 1, 0, 0, 1],
        )
        if k_nodes is None:
            k_nodes = self.model.match_parent_path(
                matmul_qk,
                ["Transpose", "Transpose", "Reshape", "Add", "MatMul"],
                [1, 0, 0, 0, 0],
            )
            if k_nodes is None:
                k_nodes = self.model.match_parent_path(
                    matmul_qk,
                    ["Transpose", "Reshape", "Add", "MatMul"],
                    [1, 0, 0, 0],
                )
                if k_nodes is None:
                    logger.debug("fuse_conformer_attention: failed to match k path")
                    return
        else:
            concat_k = k_nodes[1]
            concat_parent = self.model.get_parent(concat_k, 0, None)
            past_k = concat_parent.output[0]
            present_k = concat_k.output[0]

        add_k, matmul_k = k_nodes[-2], k_nodes[-1]

        num_heads, hidden_size = self.get_num_heads_and_hidden_size(reshape_q)
        if num_heads <= 0 or hidden_size <= 0 or (hidden_size % num_heads) != 0:
            logger.debug("fuse_conformer_attention: failed to detect num_heads or hidden_size")
            return

        new_node = None
        use_packed_attention_op = (
            matmul_q.input[0] == matmul_k.input[0] and matmul_k.input[0] == matmul_v.input[0] and extra_q_nodes is None
        )
        if use_packed_attention_op:
            # Self-attention, use Attention op
            new_node = self.create_attention_node(
                mask_index=attn_mask,
                q_matmul=matmul_q,
                k_matmul=matmul_k,
                v_matmul=matmul_v,
                q_add=add_q,
                k_add=add_k,
                v_add=add_v,
                num_heads=num_heads,
                hidden_size=hidden_size,
                first_input=matmul_q.input[0],
                output=reshape_qkv.output[0],
                add_qk_str=add_qk.input[1],
                past_k=past_k,
                past_v=past_v,
                present_k=present_k,
                present_v=present_v,
            )
        else:
            new_node = self.create_multihead_attention_node(
                q_matmul=matmul_q,
                k_matmul=matmul_k,
                v_matmul=matmul_v,
                q_add=add_q,
                k_add=add_k,
                v_add=add_v,
                num_heads=num_heads,
                hidden_size=hidden_size,
                output=reshape_qkv.output[0],
                key_padding_mask=attn_mask,
                add_qk=add_qk.input[1],
                past_k=past_k,
                past_v=past_v,
                present_k=present_k,
                present_v=present_v,
            )

        if new_node is None:
            logger.debug("fuse_conformer_attention: MultiHeadAttention node creation failed")
            return

        self.nodes_to_add.append(new_node)
        self.node_name_to_graph_name[new_node.name] = self.this_graph_name

        self.nodes_to_remove.extend([reshape_qkv, transpose_qkv, matmul_qkv])
        self.nodes_to_remove.extend(qk_nodes)

        # When using MultiHeadAttention, keep MatMul nodes unfused in original graph
        if not use_packed_attention_op:
            if q_nodes[-1].op_type == "MatMul":
                q_nodes.pop()
            if k_nodes[-1].op_type == "MatMul":
                k_nodes.pop()
            if v_nodes[-1].op_type == "MatMul":
                v_nodes.pop()

        if extra_q_nodes is None:
            # Don't remove Q nodes for conformer-transducer (CT) model since it has
            # an extra set of nodes attached to the output of the Q path that are not
            # part of the attention computation
            self.nodes_to_remove.extend(q_nodes)

        self.nodes_to_remove.extend(k_nodes)
        self.nodes_to_remove.extend(v_nodes)

        # Use prune graph to remove mask nodes since they are shared by all attention nodes.
        self.prune_graph = True

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sqlalchemy\event\api.py ===
# event/api.py
# Copyright (C) 2005-2025 the SQLAlchemy authors and contributors
# <see AUTHORS file>
#
# This module is part of SQLAlchemy and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

"""Public API functions for the event system.

"""
from __future__ import annotations

from typing import Any
from typing import Callable

from .base import _registrars
from .registry import _ET
from .registry import _EventKey
from .registry import _ListenerFnType
from .. import exc
from .. import util


CANCEL = util.symbol("CANCEL")
NO_RETVAL = util.symbol("NO_RETVAL")


def _event_key(
    target: _ET, identifier: str, fn: _ListenerFnType
) -> _EventKey[_ET]:
    for evt_cls in _registrars[identifier]:
        tgt = evt_cls._accept_with(target, identifier)
        if tgt is not None:
            return _EventKey(target, identifier, fn, tgt)
    else:
        raise exc.InvalidRequestError(
            "No such event '%s' for target '%s'" % (identifier, target)
        )


def listen(
    target: Any, identifier: str, fn: Callable[..., Any], *args: Any, **kw: Any
) -> None:
    """Register a listener function for the given target.

    The :func:`.listen` function is part of the primary interface for the
    SQLAlchemy event system, documented at :ref:`event_toplevel`.

    e.g.::

        from sqlalchemy import event
        from sqlalchemy.schema import UniqueConstraint


        def unique_constraint_name(const, table):
            const.name = "uq_%s_%s" % (table.name, list(const.columns)[0].name)


        event.listen(
            UniqueConstraint, "after_parent_attach", unique_constraint_name
        )

    :param bool insert: The default behavior for event handlers is to append
      the decorated user defined function to an internal list of registered
      event listeners upon discovery. If a user registers a function with
      ``insert=True``, SQLAlchemy will insert (prepend) the function to the
      internal list upon discovery. This feature is not typically used or
      recommended by the SQLAlchemy maintainers, but is provided to ensure
      certain user defined functions can run before others, such as when
      :ref:`Changing the sql_mode in MySQL <mysql_sql_mode>`.

    :param bool named: When using named argument passing, the names listed in
      the function argument specification will be used as keys in the
      dictionary.
      See :ref:`event_named_argument_styles`.

    :param bool once: Private/Internal API usage. Deprecated.  This parameter
      would provide that an event function would run only once per given
      target. It does not however imply automatic de-registration of the
      listener function; associating an arbitrarily high number of listeners
      without explicitly removing them will cause memory to grow unbounded even
      if ``once=True`` is specified.

    :param bool propagate: The ``propagate`` kwarg is available when working
      with ORM instrumentation and mapping events.
      See :class:`_ormevent.MapperEvents` and
      :meth:`_ormevent.MapperEvents.before_mapper_configured` for examples.

    :param bool retval: This flag applies only to specific event listeners,
      each of which includes documentation explaining when it should be used.
      By default, no listener ever requires a return value.
      However, some listeners do support special behaviors for return values,
      and include in their documentation that the ``retval=True`` flag is
      necessary for a return value to be processed.

      Event listener suites that make use of :paramref:`_event.listen.retval`
      include :class:`_events.ConnectionEvents` and
      :class:`_ormevent.AttributeEvents`.

    .. note::

        The :func:`.listen` function cannot be called at the same time
        that the target event is being run.   This has implications
        for thread safety, and also means an event cannot be added
        from inside the listener function for itself.  The list of
        events to be run are present inside of a mutable collection
        that can't be changed during iteration.

        Event registration and removal is not intended to be a "high
        velocity" operation; it is a configurational operation.  For
        systems that need to quickly associate and deassociate with
        events at high scale, use a mutable structure that is handled
        from inside of a single listener.

    .. seealso::

        :func:`.listens_for`

        :func:`.remove`

    """

    _event_key(target, identifier, fn).listen(*args, **kw)


def listens_for(
    target: Any, identifier: str, *args: Any, **kw: Any
) -> Callable[[Callable[..., Any]], Callable[..., Any]]:
    """Decorate a function as a listener for the given target + identifier.

    The :func:`.listens_for` decorator is part of the primary interface for the
    SQLAlchemy event system, documented at :ref:`event_toplevel`.

    This function generally shares the same kwargs as :func:`.listen`.

    e.g.::

        from sqlalchemy import event
        from sqlalchemy.schema import UniqueConstraint


        @event.listens_for(UniqueConstraint, "after_parent_attach")
        def unique_constraint_name(const, table):
            const.name = "uq_%s_%s" % (table.name, list(const.columns)[0].name)

    A given function can also be invoked for only the first invocation
    of the event using the ``once`` argument::

        @event.listens_for(Mapper, "before_configure", once=True)
        def on_config():
            do_config()

    .. warning:: The ``once`` argument does not imply automatic de-registration
       of the listener function after it has been invoked a first time; a
       listener entry will remain associated with the target object.
       Associating an arbitrarily high number of listeners without explicitly
       removing them will cause memory to grow unbounded even if ``once=True``
       is specified.

    .. seealso::

        :func:`.listen` - general description of event listening

    """

    def decorate(fn: Callable[..., Any]) -> Callable[..., Any]:
        listen(target, identifier, fn, *args, **kw)
        return fn

    return decorate


def remove(target: Any, identifier: str, fn: Callable[..., Any]) -> None:
    """Remove an event listener.

    The arguments here should match exactly those which were sent to
    :func:`.listen`; all the event registration which proceeded as a result
    of this call will be reverted by calling :func:`.remove` with the same
    arguments.

    e.g.::

        # if a function was registered like this...
        @event.listens_for(SomeMappedClass, "before_insert", propagate=True)
        def my_listener_function(*arg):
            pass


        # ... it's removed like this
        event.remove(SomeMappedClass, "before_insert", my_listener_function)

    Above, the listener function associated with ``SomeMappedClass`` was also
    propagated to subclasses of ``SomeMappedClass``; the :func:`.remove`
    function will revert all of these operations.

    .. note::

        The :func:`.remove` function cannot be called at the same time
        that the target event is being run.   This has implications
        for thread safety, and also means an event cannot be removed
        from inside the listener function for itself.  The list of
        events to be run are present inside of a mutable collection
        that can't be changed during iteration.

        Event registration and removal is not intended to be a "high
        velocity" operation; it is a configurational operation.  For
        systems that need to quickly associate and deassociate with
        events at high scale, use a mutable structure that is handled
        from inside of a single listener.

    .. seealso::

        :func:`.listen`

    """
    _event_key(target, identifier, fn).remove()


def contains(target: Any, identifier: str, fn: Callable[..., Any]) -> bool:
    """Return True if the given target/ident/fn is set up to listen."""

    return _event_key(target, identifier, fn).contains()

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\concrete\gosper.py ===
"""Gosper's algorithm for hypergeometric summation. """

from sympy.core import S, Dummy, symbols
from sympy.polys import Poly, parallel_poly_from_expr, factor
from sympy.utilities.iterables import is_sequence


def gosper_normal(f, g, n, polys=True):
    r"""
    Compute the Gosper's normal form of ``f`` and ``g``.

    Explanation
    ===========

    Given relatively prime univariate polynomials ``f`` and ``g``,
    rewrite their quotient to a normal form defined as follows:

    .. math::
        \frac{f(n)}{g(n)} = Z \cdot \frac{A(n) C(n+1)}{B(n) C(n)}

    where ``Z`` is an arbitrary constant and ``A``, ``B``, ``C`` are
    monic polynomials in ``n`` with the following properties:

    1. `\gcd(A(n), B(n+h)) = 1 \forall h \in \mathbb{N}`
    2. `\gcd(B(n), C(n+1)) = 1`
    3. `\gcd(A(n), C(n)) = 1`

    This normal form, or rational factorization in other words, is a
    crucial step in Gosper's algorithm and in solving of difference
    equations. It can be also used to decide if two hypergeometric
    terms are similar or not.

    This procedure will return a tuple containing elements of this
    factorization in the form ``(Z*A, B, C)``.

    Examples
    ========

    >>> from sympy.concrete.gosper import gosper_normal
    >>> from sympy.abc import n

    >>> gosper_normal(4*n+5, 2*(4*n+1)*(2*n+3), n, polys=False)
    (1/4, n + 3/2, n + 1/4)

    """
    (p, q), opt = parallel_poly_from_expr(
        (f, g), n, field=True, extension=True)

    a, A = p.LC(), p.monic()
    b, B = q.LC(), q.monic()

    C, Z = A.one, a/b
    h = Dummy('h')

    D = Poly(n + h, n, h, domain=opt.domain)

    R = A.resultant(B.compose(D))
    roots = {r for r in R.ground_roots().keys() if r.is_Integer and r >= 0}
    for i in sorted(roots):
        d = A.gcd(B.shift(+i))

        A = A.quo(d)
        B = B.quo(d.shift(-i))

        for j in range(1, i + 1):
            C *= d.shift(-j)

    A = A.mul_ground(Z)

    if not polys:
        A = A.as_expr()
        B = B.as_expr()
        C = C.as_expr()

    return A, B, C


def gosper_term(f, n):
    r"""
    Compute Gosper's hypergeometric term for ``f``.

    Explanation
    ===========

    Suppose ``f`` is a hypergeometric term such that:

    .. math::
        s_n = \sum_{k=0}^{n-1} f_k

    and `f_k` does not depend on `n`. Returns a hypergeometric
    term `g_n` such that `g_{n+1} - g_n = f_n`.

    Examples
    ========

    >>> from sympy.concrete.gosper import gosper_term
    >>> from sympy import factorial
    >>> from sympy.abc import n

    >>> gosper_term((4*n + 1)*factorial(n)/factorial(2*n + 1), n)
    (-n - 1/2)/(n + 1/4)

    """
    from sympy.simplify import hypersimp
    r = hypersimp(f, n)

    if r is None:
        return None    # 'f' is *not* a hypergeometric term

    p, q = r.as_numer_denom()

    A, B, C = gosper_normal(p, q, n)
    B = B.shift(-1)

    N = S(A.degree())
    M = S(B.degree())
    K = S(C.degree())

    if (N != M) or (A.LC() != B.LC()):
        D = {K - max(N, M)}
    elif not N:
        D = {K - N + 1, S.Zero}
    else:
        D = {K - N + 1, (B.nth(N - 1) - A.nth(N - 1))/A.LC()}

    for d in set(D):
        if not d.is_Integer or d < 0:
            D.remove(d)

    if not D:
        return None    # 'f(n)' is *not* Gosper-summable

    d = max(D)

    coeffs = symbols('c:%s' % (d + 1), cls=Dummy)
    domain = A.get_domain().inject(*coeffs)

    x = Poly(coeffs, n, domain=domain)
    H = A*x.shift(1) - B*x - C

    from sympy.solvers.solvers import solve
    solution = solve(H.coeffs(), coeffs)

    if solution is None:
        return None    # 'f(n)' is *not* Gosper-summable

    x = x.as_expr().subs(solution)

    for coeff in coeffs:
        if coeff not in solution:
            x = x.subs(coeff, 0)

    if x.is_zero:
        return None    # 'f(n)' is *not* Gosper-summable
    else:
        return B.as_expr()*x/C.as_expr()


def gosper_sum(f, k):
    r"""
    Gosper's hypergeometric summation algorithm.

    Explanation
    ===========

    Given a hypergeometric term ``f`` such that:

    .. math ::
        s_n = \sum_{k=0}^{n-1} f_k

    and `f(n)` does not depend on `n`, returns `g_{n} - g(0)` where
    `g_{n+1} - g_n = f_n`, or ``None`` if `s_n` cannot be expressed
    in closed form as a sum of hypergeometric terms.

    Examples
    ========

    >>> from sympy.concrete.gosper import gosper_sum
    >>> from sympy import factorial
    >>> from sympy.abc import n, k

    >>> f = (4*k + 1)*factorial(k)/factorial(2*k + 1)
    >>> gosper_sum(f, (k, 0, n))
    (-factorial(n) + 2*factorial(2*n + 1))/factorial(2*n + 1)
    >>> _.subs(n, 2) == sum(f.subs(k, i) for i in [0, 1, 2])
    True
    >>> gosper_sum(f, (k, 3, n))
    (-60*factorial(n) + factorial(2*n + 1))/(60*factorial(2*n + 1))
    >>> _.subs(n, 5) == sum(f.subs(k, i) for i in [3, 4, 5])
    True

    References
    ==========

    .. [1] Marko Petkovsek, Herbert S. Wilf, Doron Zeilberger, A = B,
           AK Peters, Ltd., Wellesley, MA, USA, 1997, pp. 73--100

    """
    indefinite = False

    if is_sequence(k):
        k, a, b = k
    else:
        indefinite = True

    g = gosper_term(f, k)

    if g is None:
        return None

    if indefinite:
        result = f*g
    else:
        result = (f*(g + 1)).subs(k, b) - (f*g).subs(k, a)

        if result is S.NaN:
            try:
                result = (f*(g + 1)).limit(k, b) - (f*g).limit(k, a)
            except NotImplementedError:
                result = None

    return factor(result)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\sympy\simplify\ratsimp.py ===
from itertools import combinations_with_replacement
from sympy.core import symbols, Add, Dummy
from sympy.core.numbers import Rational
from sympy.polys import cancel, ComputationFailed, parallel_poly_from_expr, reduced, Poly
from sympy.polys.monomials import Monomial, monomial_div
from sympy.polys.polyerrors import DomainError, PolificationFailed
from sympy.utilities.misc import debug, debugf

def ratsimp(expr):
    """
    Put an expression over a common denominator, cancel and reduce.

    Examples
    ========

    >>> from sympy import ratsimp
    >>> from sympy.abc import x, y
    >>> ratsimp(1/x + 1/y)
    (x + y)/(x*y)
    """

    f, g = cancel(expr).as_numer_denom()
    try:
        Q, r = reduced(f, [g], field=True, expand=False)
    except ComputationFailed:
        return f/g

    return Add(*Q) + cancel(r/g)


def ratsimpmodprime(expr, G, *gens, quick=True, polynomial=False, **args):
    """
    Simplifies a rational expression ``expr`` modulo the prime ideal
    generated by ``G``.  ``G`` should be a Groebner basis of the
    ideal.

    Examples
    ========

    >>> from sympy.simplify.ratsimp import ratsimpmodprime
    >>> from sympy.abc import x, y
    >>> eq = (x + y**5 + y)/(x - y)
    >>> ratsimpmodprime(eq, [x*y**5 - x - y], x, y, order='lex')
    (-x**2 - x*y - x - y)/(-x**2 + x*y)

    If ``polynomial`` is ``False``, the algorithm computes a rational
    simplification which minimizes the sum of the total degrees of
    the numerator and the denominator.

    If ``polynomial`` is ``True``, this function just brings numerator and
    denominator into a canonical form. This is much faster, but has
    potentially worse results.

    References
    ==========

    .. [1] M. Monagan, R. Pearce, Rational Simplification Modulo a Polynomial
        Ideal, https://dl.acm.org/doi/pdf/10.1145/1145768.1145809
        (specifically, the second algorithm)
    """
    from sympy.solvers.solvers import solve

    debug('ratsimpmodprime', expr)

    # usual preparation of polynomials:

    num, denom = cancel(expr).as_numer_denom()

    try:
        polys, opt = parallel_poly_from_expr([num, denom] + G, *gens, **args)
    except PolificationFailed:
        return expr

    domain = opt.domain

    if domain.has_assoc_Field:
        opt.domain = domain.get_field()
    else:
        raise DomainError(
            "Cannot compute rational simplification over %s" % domain)

    # compute only once
    leading_monomials = [g.LM(opt.order) for g in polys[2:]]
    tested = set()

    def staircase(n):
        """
        Compute all monomials with degree less than ``n`` that are
        not divisible by any element of ``leading_monomials``.
        """
        if n == 0:
            return [1]
        S = []
        for mi in combinations_with_replacement(range(len(opt.gens)), n):
            m = [0]*len(opt.gens)
            for i in mi:
                m[i] += 1
            if all(monomial_div(m, lmg) is None for lmg in
                   leading_monomials):
                S.append(m)

        return [Monomial(s).as_expr(*opt.gens) for s in S] + staircase(n - 1)

    def _ratsimpmodprime(a, b, allsol, N=0, D=0):
        r"""
        Computes a rational simplification of ``a/b`` which minimizes
        the sum of the total degrees of the numerator and the denominator.

        Explanation
        ===========

        The algorithm proceeds by looking at ``a * d - b * c`` modulo
        the ideal generated by ``G`` for some ``c`` and ``d`` with degree
        less than ``a`` and ``b`` respectively.
        The coefficients of ``c`` and ``d`` are indeterminates and thus
        the coefficients of the normalform of ``a * d - b * c`` are
        linear polynomials in these indeterminates.
        If these linear polynomials, considered as system of
        equations, have a nontrivial solution, then `\frac{a}{b}
        \equiv \frac{c}{d}` modulo the ideal generated by ``G``. So,
        by construction, the degree of ``c`` and ``d`` is less than
        the degree of ``a`` and ``b``, so a simpler representation
        has been found.
        After a simpler representation has been found, the algorithm
        tries to reduce the degree of the numerator and denominator
        and returns the result afterwards.

        As an extension, if quick=False, we look at all possible degrees such
        that the total degree is less than *or equal to* the best current
        solution. We retain a list of all solutions of minimal degree, and try
        to find the best one at the end.
        """
        c, d = a, b
        steps = 0

        maxdeg = a.total_degree() + b.total_degree()
        if quick:
            bound = maxdeg - 1
        else:
            bound = maxdeg
        while N + D <= bound:
            if (N, D) in tested:
                break
            tested.add((N, D))

            M1 = staircase(N)
            M2 = staircase(D)
            debugf('%s / %s: %s, %s', (N, D, M1, M2))

            Cs = symbols("c:%d" % len(M1), cls=Dummy)
            Ds = symbols("d:%d" % len(M2), cls=Dummy)
            ng = Cs + Ds

            c_hat = Poly(
                sum(Cs[i] * M1[i] for i in range(len(M1))), opt.gens + ng)
            d_hat = Poly(
                sum(Ds[i] * M2[i] for i in range(len(M2))), opt.gens + ng)

            r = reduced(a * d_hat - b * c_hat, G, opt.gens + ng,
                        order=opt.order, polys=True)[1]

            S = Poly(r, gens=opt.gens).coeffs()
            sol = solve(S, Cs + Ds, particular=True, quick=True)

            if sol and not all(s == 0 for s in sol.values()):
                c = c_hat.subs(sol)
                d = d_hat.subs(sol)

                # The "free" variables occurring before as parameters
                # might still be in the substituted c, d, so set them
                # to the value chosen before:
                c = c.subs(dict(list(zip(Cs + Ds, [1] * (len(Cs) + len(Ds))))))
                d = d.subs(dict(list(zip(Cs + Ds, [1] * (len(Cs) + len(Ds))))))

                c = Poly(c, opt.gens)
                d = Poly(d, opt.gens)
                if d == 0:
                    raise ValueError('Ideal not prime?')

                allsol.append((c_hat, d_hat, S, Cs + Ds))
                if N + D != maxdeg:
                    allsol = [allsol[-1]]

                break

            steps += 1
            N += 1
            D += 1

        if steps > 0:
            c, d, allsol = _ratsimpmodprime(c, d, allsol, N, D - steps)
            c, d, allsol = _ratsimpmodprime(c, d, allsol, N - steps, D)

        return c, d, allsol

    # preprocessing. this improves performance a bit when deg(num)
    # and deg(denom) are large:
    num = reduced(num, G, opt.gens, order=opt.order)[1]
    denom = reduced(denom, G, opt.gens, order=opt.order)[1]

    if polynomial:
        return (num/denom).cancel()

    c, d, allsol = _ratsimpmodprime(
        Poly(num, opt.gens, domain=opt.domain), Poly(denom, opt.gens, domain=opt.domain), [])
    if not quick and allsol:
        debugf('Looking for best minimal solution. Got: %s', len(allsol))
        newsol = []
        for c_hat, d_hat, S, ng in allsol:
            sol = solve(S, ng, particular=True, quick=False)
            # all values of sol should be numbers; if not, solve is broken
            newsol.append((c_hat.subs(sol), d_hat.subs(sol)))
        c, d = min(newsol, key=lambda x: len(x[0].terms()) + len(x[1].terms()))

    if not domain.is_Field:
        cn, c = c.clear_denoms(convert=True)
        dn, d = d.clear_denoms(convert=True)
        r = Rational(cn, dn)
    else:
        r = Rational(1)

    return (c*r.q)/(d*r.p)

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\pip\_vendor\urllib3\util\ssltransport.py ===
import io
import socket
import ssl

from ..exceptions import ProxySchemeUnsupported
from ..packages import six

SSL_BLOCKSIZE = 16384


class SSLTransport:
    """
    The SSLTransport wraps an existing socket and establishes an SSL connection.

    Contrary to Python's implementation of SSLSocket, it allows you to chain
    multiple TLS connections together. It's particularly useful if you need to
    implement TLS within TLS.

    The class supports most of the socket API operations.
    """

    @staticmethod
    def _validate_ssl_context_for_tls_in_tls(ssl_context):
        """
        Raises a ProxySchemeUnsupported if the provided ssl_context can't be used
        for TLS in TLS.

        The only requirement is that the ssl_context provides the 'wrap_bio'
        methods.
        """

        if not hasattr(ssl_context, "wrap_bio"):
            if six.PY2:
                raise ProxySchemeUnsupported(
                    "TLS in TLS requires SSLContext.wrap_bio() which isn't "
                    "supported on Python 2"
                )
            else:
                raise ProxySchemeUnsupported(
                    "TLS in TLS requires SSLContext.wrap_bio() which isn't "
                    "available on non-native SSLContext"
                )

    def __init__(
        self, socket, ssl_context, server_hostname=None, suppress_ragged_eofs=True
    ):
        """
        Create an SSLTransport around socket using the provided ssl_context.
        """
        self.incoming = ssl.MemoryBIO()
        self.outgoing = ssl.MemoryBIO()

        self.suppress_ragged_eofs = suppress_ragged_eofs
        self.socket = socket

        self.sslobj = ssl_context.wrap_bio(
            self.incoming, self.outgoing, server_hostname=server_hostname
        )

        # Perform initial handshake.
        self._ssl_io_loop(self.sslobj.do_handshake)

    def __enter__(self):
        return self

    def __exit__(self, *_):
        self.close()

    def fileno(self):
        return self.socket.fileno()

    def read(self, len=1024, buffer=None):
        return self._wrap_ssl_read(len, buffer)

    def recv(self, len=1024, flags=0):
        if flags != 0:
            raise ValueError("non-zero flags not allowed in calls to recv")
        return self._wrap_ssl_read(len)

    def recv_into(self, buffer, nbytes=None, flags=0):
        if flags != 0:
            raise ValueError("non-zero flags not allowed in calls to recv_into")
        if buffer and (nbytes is None):
            nbytes = len(buffer)
        elif nbytes is None:
            nbytes = 1024
        return self.read(nbytes, buffer)

    def sendall(self, data, flags=0):
        if flags != 0:
            raise ValueError("non-zero flags not allowed in calls to sendall")
        count = 0
        with memoryview(data) as view, view.cast("B") as byte_view:
            amount = len(byte_view)
            while count < amount:
                v = self.send(byte_view[count:])
                count += v

    def send(self, data, flags=0):
        if flags != 0:
            raise ValueError("non-zero flags not allowed in calls to send")
        response = self._ssl_io_loop(self.sslobj.write, data)
        return response

    def makefile(
        self, mode="r", buffering=None, encoding=None, errors=None, newline=None
    ):
        """
        Python's httpclient uses makefile and buffered io when reading HTTP
        messages and we need to support it.

        This is unfortunately a copy and paste of socket.py makefile with small
        changes to point to the socket directly.
        """
        if not set(mode) <= {"r", "w", "b"}:
            raise ValueError("invalid mode %r (only r, w, b allowed)" % (mode,))

        writing = "w" in mode
        reading = "r" in mode or not writing
        assert reading or writing
        binary = "b" in mode
        rawmode = ""
        if reading:
            rawmode += "r"
        if writing:
            rawmode += "w"
        raw = socket.SocketIO(self, rawmode)
        self.socket._io_refs += 1
        if buffering is None:
            buffering = -1
        if buffering < 0:
            buffering = io.DEFAULT_BUFFER_SIZE
        if buffering == 0:
            if not binary:
                raise ValueError("unbuffered streams must be binary")
            return raw
        if reading and writing:
            buffer = io.BufferedRWPair(raw, raw, buffering)
        elif reading:
            buffer = io.BufferedReader(raw, buffering)
        else:
            assert writing
            buffer = io.BufferedWriter(raw, buffering)
        if binary:
            return buffer
        text = io.TextIOWrapper(buffer, encoding, errors, newline)
        text.mode = mode
        return text

    def unwrap(self):
        self._ssl_io_loop(self.sslobj.unwrap)

    def close(self):
        self.socket.close()

    def getpeercert(self, binary_form=False):
        return self.sslobj.getpeercert(binary_form)

    def version(self):
        return self.sslobj.version()

    def cipher(self):
        return self.sslobj.cipher()

    def selected_alpn_protocol(self):
        return self.sslobj.selected_alpn_protocol()

    def selected_npn_protocol(self):
        return self.sslobj.selected_npn_protocol()

    def shared_ciphers(self):
        return self.sslobj.shared_ciphers()

    def compression(self):
        return self.sslobj.compression()

    def settimeout(self, value):
        self.socket.settimeout(value)

    def gettimeout(self):
        return self.socket.gettimeout()

    def _decref_socketios(self):
        self.socket._decref_socketios()

    def _wrap_ssl_read(self, len, buffer=None):
        try:
            return self._ssl_io_loop(self.sslobj.read, len, buffer)
        except ssl.SSLError as e:
            if e.errno == ssl.SSL_ERROR_EOF and self.suppress_ragged_eofs:
                return 0  # eof, return 0.
            else:
                raise

    def _ssl_io_loop(self, func, *args):
        """Performs an I/O loop between incoming/outgoing and the socket."""
        should_loop = True
        ret = None

        while should_loop:
            errno = None
            try:
                ret = func(*args)
            except ssl.SSLError as e:
                if e.errno not in (ssl.SSL_ERROR_WANT_READ, ssl.SSL_ERROR_WANT_WRITE):
                    # WANT_READ, and WANT_WRITE are expected, others are not.
                    raise e
                errno = e.errno

            buf = self.outgoing.read()
            self.socket.sendall(buf)

            if errno is None:
                should_loop = False
            elif errno == ssl.SSL_ERROR_WANT_READ:
                buf = self.socket.recv(SSL_BLOCKSIZE)
                if buf:
                    self.incoming.write(buf)
                else:
                    self.incoming.write_eof()
        return ret

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\virtual_authenticator.py ===
# Licensed to the Software Freedom Conservancy (SFC) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The SFC licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

import functools
from base64 import urlsafe_b64decode, urlsafe_b64encode
from enum import Enum
from typing import Any, Dict, Optional, Union


class Protocol(str, Enum):
    """Protocol to communicate with the authenticator."""

    CTAP2: str = "ctap2"
    U2F: str = "ctap1/u2f"


class Transport(str, Enum):
    """Transport method to communicate with the authenticator."""

    BLE: str = "ble"
    USB: str = "usb"
    NFC: str = "nfc"
    INTERNAL: str = "internal"


class VirtualAuthenticatorOptions:
    # These are so unnecessary but are now public API so we can't remove them without deprecating first.
    # These should not be class level state in here.
    Protocol = Protocol
    Transport = Transport

    def __init__(
        self,
        protocol: str = Protocol.CTAP2,
        transport: str = Transport.USB,
        has_resident_key: bool = False,
        has_user_verification: bool = False,
        is_user_consenting: bool = True,
        is_user_verified: bool = False,
    ) -> None:
        """Constructor.

        Initialize VirtualAuthenticatorOptions object.
        """

        self.protocol: str = protocol
        self.transport: str = transport
        self.has_resident_key: bool = has_resident_key
        self.has_user_verification: bool = has_user_verification
        self.is_user_consenting: bool = is_user_consenting
        self.is_user_verified: bool = is_user_verified

    def to_dict(self) -> Dict[str, Union[str, bool]]:
        return {
            "protocol": self.protocol,
            "transport": self.transport,
            "hasResidentKey": self.has_resident_key,
            "hasUserVerification": self.has_user_verification,
            "isUserConsenting": self.is_user_consenting,
            "isUserVerified": self.is_user_verified,
        }


class Credential:
    def __init__(
        self,
        credential_id: bytes,
        is_resident_credential: bool,
        rp_id: str,
        user_handle: Optional[bytes],
        private_key: bytes,
        sign_count: int,
    ):
        """Constructor. A credential stored in a virtual authenticator.
        https://w3c.github.io/webauthn/#credential-parameters.

        :Args:
            - credential_id (bytes): Unique base64 encoded string.
            - is_resident_credential (bool): Whether the credential is client-side discoverable.
            - rp_id (str): Relying party identifier.
            - user_handle (bytes): userHandle associated to the credential. Must be Base64 encoded string. Can be None.
            - private_key (bytes): Base64 encoded PKCS#8 private key.
            - sign_count (int): intital value for a signature counter.
        """
        self._id = credential_id
        self._is_resident_credential = is_resident_credential
        self._rp_id = rp_id
        self._user_handle = user_handle
        self._private_key = private_key
        self._sign_count = sign_count

    @property
    def id(self) -> str:
        return urlsafe_b64encode(self._id).decode()

    @property
    def is_resident_credential(self) -> bool:
        return self._is_resident_credential

    @property
    def rp_id(self) -> str:
        return self._rp_id

    @property
    def user_handle(self) -> Optional[str]:
        if self._user_handle:
            return urlsafe_b64encode(self._user_handle).decode()
        return None

    @property
    def private_key(self) -> str:
        return urlsafe_b64encode(self._private_key).decode()

    @property
    def sign_count(self) -> int:
        return self._sign_count

    @classmethod
    def create_non_resident_credential(cls, id: bytes, rp_id: str, private_key: bytes, sign_count: int) -> "Credential":
        """Creates a non-resident (i.e. stateless) credential.

        :Args:
          - id (bytes): Unique base64 encoded string.
          - rp_id (str): Relying party identifier.
          - private_key (bytes): Base64 encoded PKCS
          - sign_count (int): intital value for a signature counter.

        :Returns:
          - Credential: A non-resident credential.
        """
        return cls(id, False, rp_id, None, private_key, sign_count)

    @classmethod
    def create_resident_credential(
        cls, id: bytes, rp_id: str, user_handle: Optional[bytes], private_key: bytes, sign_count: int
    ) -> "Credential":
        """Creates a resident (i.e. stateful) credential.

        :Args:
          - id (bytes): Unique base64 encoded string.
          - rp_id (str): Relying party identifier.
          - user_handle (bytes): userHandle associated to the credential. Must be Base64 encoded string.
          - private_key (bytes): Base64 encoded PKCS
          - sign_count (int): intital value for a signature counter.

        :returns:
          - Credential: A resident credential.
        """
        return cls(id, True, rp_id, user_handle, private_key, sign_count)

    def to_dict(self) -> Dict[str, Any]:
        credential_data = {
            "credentialId": self.id,
            "isResidentCredential": self._is_resident_credential,
            "rpId": self.rp_id,
            "privateKey": self.private_key,
            "signCount": self.sign_count,
        }

        if self.user_handle:
            credential_data["userHandle"] = self.user_handle

        return credential_data

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "Credential":
        _id = urlsafe_b64decode(f"{data['credentialId']}==")
        is_resident_credential = bool(data["isResidentCredential"])
        rp_id = data.get("rpId", None)
        private_key = urlsafe_b64decode(f"{data['privateKey']}==")
        sign_count = int(data["signCount"])
        user_handle = urlsafe_b64decode(f"{data['userHandle']}==") if data.get("userHandle", None) else None

        return cls(_id, is_resident_credential, rp_id, user_handle, private_key, sign_count)

    def __str__(self) -> str:
        return f"Credential(id={self.id}, is_resident_credential={self.is_resident_credential}, rp_id={self.rp_id},\
            user_handle={self.user_handle}, private_key={self.private_key}, sign_count={self.sign_count})"


def required_chromium_based_browser(func):
    """A decorator to ensure that the client used is a chromium based
    browser."""

    @functools.wraps(func)
    def wrapper(self, *args, **kwargs):
        assert self.caps["browserName"].lower() not in [
            "firefox",
            "safari",
        ], "This only currently works in Chromium based browsers"
        return func(self, *args, **kwargs)

    return wrapper


def required_virtual_authenticator(func):
    """A decorator to ensure that the function is called with a virtual
    authenticator."""

    @functools.wraps(func)
    @required_chromium_based_browser
    def wrapper(self, *args, **kwargs):
        if not self.virtual_authenticator_id:
            raise ValueError("This function requires a virtual authenticator to be set.")
        return func(self, *args, **kwargs)

    return wrapper

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v135\dom_storage.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: DOMStorage (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

class SerializedStorageKey(str):
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> SerializedStorageKey:
        return cls(json)

    def __repr__(self):
        return 'SerializedStorageKey({})'.format(super().__repr__())


@dataclass
class StorageId:
    '''
    DOM Storage identifier.
    '''
    #: Whether the storage is local storage (not session storage).
    is_local_storage: bool

    #: Security origin for the storage.
    security_origin: typing.Optional[str] = None

    #: Represents a key by which DOM Storage keys its CachedStorageAreas
    storage_key: typing.Optional[SerializedStorageKey] = None

    def to_json(self):
        json = dict()
        json['isLocalStorage'] = self.is_local_storage
        if self.security_origin is not None:
            json['securityOrigin'] = self.security_origin
        if self.storage_key is not None:
            json['storageKey'] = self.storage_key.to_json()
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            is_local_storage=bool(json['isLocalStorage']),
            security_origin=str(json['securityOrigin']) if 'securityOrigin' in json else None,
            storage_key=SerializedStorageKey.from_json(json['storageKey']) if 'storageKey' in json else None,
        )


class Item(list):
    '''
    DOM Storage item.
    '''
    def to_json(self) -> typing.List[str]:
        return self

    @classmethod
    def from_json(cls, json: typing.List[str]) -> Item:
        return cls(json)

    def __repr__(self):
        return 'Item({})'.format(super().__repr__())


def clear(
        storage_id: StorageId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    :param storage_id:
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.clear',
        'params': params,
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disables storage tracking, prevents storage events from being sent to the client.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.disable',
    }
    json = yield cmd_dict


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enables storage tracking, storage events will now be delivered to the client.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.enable',
    }
    json = yield cmd_dict


def get_dom_storage_items(
        storage_id: StorageId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[Item]]:
    '''
    :param storage_id:
    :returns: 
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.getDOMStorageItems',
        'params': params,
    }
    json = yield cmd_dict
    return [Item.from_json(i) for i in json['entries']]


def remove_dom_storage_item(
        storage_id: StorageId,
        key: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    :param storage_id:
    :param key:
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    params['key'] = key
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.removeDOMStorageItem',
        'params': params,
    }
    json = yield cmd_dict


def set_dom_storage_item(
        storage_id: StorageId,
        key: str,
        value: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    :param storage_id:
    :param key:
    :param value:
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    params['key'] = key
    params['value'] = value
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.setDOMStorageItem',
        'params': params,
    }
    json = yield cmd_dict


@event_class('DOMStorage.domStorageItemAdded')
@dataclass
class DomStorageItemAdded:
    storage_id: StorageId
    key: str
    new_value: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemAdded:
        return cls(
            storage_id=StorageId.from_json(json['storageId']),
            key=str(json['key']),
            new_value=str(json['newValue'])
        )


@event_class('DOMStorage.domStorageItemRemoved')
@dataclass
class DomStorageItemRemoved:
    storage_id: StorageId
    key: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemRemoved:
        return cls(
            storage_id=StorageId.from_json(json['storageId']),
            key=str(json['key'])
        )


@event_class('DOMStorage.domStorageItemUpdated')
@dataclass
class DomStorageItemUpdated:
    storage_id: StorageId
    key: str
    old_value: str
    new_value: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemUpdated:
        return cls(
            storage_id=StorageId.from_json(json['storageId']),
            key=str(json['key']),
            old_value=str(json['oldValue']),
            new_value=str(json['newValue'])
        )


@event_class('DOMStorage.domStorageItemsCleared')
@dataclass
class DomStorageItemsCleared:
    storage_id: StorageId

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemsCleared:
        return cls(
            storage_id=StorageId.from_json(json['storageId'])
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v136\dom_storage.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: DOMStorage (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

class SerializedStorageKey(str):
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> SerializedStorageKey:
        return cls(json)

    def __repr__(self):
        return 'SerializedStorageKey({})'.format(super().__repr__())


@dataclass
class StorageId:
    '''
    DOM Storage identifier.
    '''
    #: Whether the storage is local storage (not session storage).
    is_local_storage: bool

    #: Security origin for the storage.
    security_origin: typing.Optional[str] = None

    #: Represents a key by which DOM Storage keys its CachedStorageAreas
    storage_key: typing.Optional[SerializedStorageKey] = None

    def to_json(self):
        json = dict()
        json['isLocalStorage'] = self.is_local_storage
        if self.security_origin is not None:
            json['securityOrigin'] = self.security_origin
        if self.storage_key is not None:
            json['storageKey'] = self.storage_key.to_json()
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            is_local_storage=bool(json['isLocalStorage']),
            security_origin=str(json['securityOrigin']) if 'securityOrigin' in json else None,
            storage_key=SerializedStorageKey.from_json(json['storageKey']) if 'storageKey' in json else None,
        )


class Item(list):
    '''
    DOM Storage item.
    '''
    def to_json(self) -> typing.List[str]:
        return self

    @classmethod
    def from_json(cls, json: typing.List[str]) -> Item:
        return cls(json)

    def __repr__(self):
        return 'Item({})'.format(super().__repr__())


def clear(
        storage_id: StorageId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    :param storage_id:
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.clear',
        'params': params,
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disables storage tracking, prevents storage events from being sent to the client.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.disable',
    }
    json = yield cmd_dict


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enables storage tracking, storage events will now be delivered to the client.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.enable',
    }
    json = yield cmd_dict


def get_dom_storage_items(
        storage_id: StorageId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[Item]]:
    '''
    :param storage_id:
    :returns: 
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.getDOMStorageItems',
        'params': params,
    }
    json = yield cmd_dict
    return [Item.from_json(i) for i in json['entries']]


def remove_dom_storage_item(
        storage_id: StorageId,
        key: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    :param storage_id:
    :param key:
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    params['key'] = key
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.removeDOMStorageItem',
        'params': params,
    }
    json = yield cmd_dict


def set_dom_storage_item(
        storage_id: StorageId,
        key: str,
        value: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    :param storage_id:
    :param key:
    :param value:
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    params['key'] = key
    params['value'] = value
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.setDOMStorageItem',
        'params': params,
    }
    json = yield cmd_dict


@event_class('DOMStorage.domStorageItemAdded')
@dataclass
class DomStorageItemAdded:
    storage_id: StorageId
    key: str
    new_value: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemAdded:
        return cls(
            storage_id=StorageId.from_json(json['storageId']),
            key=str(json['key']),
            new_value=str(json['newValue'])
        )


@event_class('DOMStorage.domStorageItemRemoved')
@dataclass
class DomStorageItemRemoved:
    storage_id: StorageId
    key: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemRemoved:
        return cls(
            storage_id=StorageId.from_json(json['storageId']),
            key=str(json['key'])
        )


@event_class('DOMStorage.domStorageItemUpdated')
@dataclass
class DomStorageItemUpdated:
    storage_id: StorageId
    key: str
    old_value: str
    new_value: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemUpdated:
        return cls(
            storage_id=StorageId.from_json(json['storageId']),
            key=str(json['key']),
            old_value=str(json['oldValue']),
            new_value=str(json['newValue'])
        )


@event_class('DOMStorage.domStorageItemsCleared')
@dataclass
class DomStorageItemsCleared:
    storage_id: StorageId

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemsCleared:
        return cls(
            storage_id=StorageId.from_json(json['storageId'])
        )

# === atelier-kyo-manager/.venv_backup\Lib\site-packages\selenium\webdriver\common\devtools\v137\dom_storage.py ===
# DO NOT EDIT THIS FILE!
#
# This file is generated from the CDP specification. If you need to make
# changes, edit the generator and regenerate all of the modules.
#
# CDP domain: DOMStorage (experimental)
from __future__ import annotations
from .util import event_class, T_JSON_DICT
from dataclasses import dataclass
import enum
import typing

class SerializedStorageKey(str):
    def to_json(self) -> str:
        return self

    @classmethod
    def from_json(cls, json: str) -> SerializedStorageKey:
        return cls(json)

    def __repr__(self):
        return 'SerializedStorageKey({})'.format(super().__repr__())


@dataclass
class StorageId:
    '''
    DOM Storage identifier.
    '''
    #: Whether the storage is local storage (not session storage).
    is_local_storage: bool

    #: Security origin for the storage.
    security_origin: typing.Optional[str] = None

    #: Represents a key by which DOM Storage keys its CachedStorageAreas
    storage_key: typing.Optional[SerializedStorageKey] = None

    def to_json(self):
        json = dict()
        json['isLocalStorage'] = self.is_local_storage
        if self.security_origin is not None:
            json['securityOrigin'] = self.security_origin
        if self.storage_key is not None:
            json['storageKey'] = self.storage_key.to_json()
        return json

    @classmethod
    def from_json(cls, json):
        return cls(
            is_local_storage=bool(json['isLocalStorage']),
            security_origin=str(json['securityOrigin']) if 'securityOrigin' in json else None,
            storage_key=SerializedStorageKey.from_json(json['storageKey']) if 'storageKey' in json else None,
        )


class Item(list):
    '''
    DOM Storage item.
    '''
    def to_json(self) -> typing.List[str]:
        return self

    @classmethod
    def from_json(cls, json: typing.List[str]) -> Item:
        return cls(json)

    def __repr__(self):
        return 'Item({})'.format(super().__repr__())


def clear(
        storage_id: StorageId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    :param storage_id:
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.clear',
        'params': params,
    }
    json = yield cmd_dict


def disable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Disables storage tracking, prevents storage events from being sent to the client.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.disable',
    }
    json = yield cmd_dict


def enable() -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    Enables storage tracking, storage events will now be delivered to the client.
    '''
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.enable',
    }
    json = yield cmd_dict


def get_dom_storage_items(
        storage_id: StorageId
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,typing.List[Item]]:
    '''
    :param storage_id:
    :returns: 
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.getDOMStorageItems',
        'params': params,
    }
    json = yield cmd_dict
    return [Item.from_json(i) for i in json['entries']]


def remove_dom_storage_item(
        storage_id: StorageId,
        key: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    :param storage_id:
    :param key:
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    params['key'] = key
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.removeDOMStorageItem',
        'params': params,
    }
    json = yield cmd_dict


def set_dom_storage_item(
        storage_id: StorageId,
        key: str,
        value: str
    ) -> typing.Generator[T_JSON_DICT,T_JSON_DICT,None]:
    '''
    :param storage_id:
    :param key:
    :param value:
    '''
    params: T_JSON_DICT = dict()
    params['storageId'] = storage_id.to_json()
    params['key'] = key
    params['value'] = value
    cmd_dict: T_JSON_DICT = {
        'method': 'DOMStorage.setDOMStorageItem',
        'params': params,
    }
    json = yield cmd_dict


@event_class('DOMStorage.domStorageItemAdded')
@dataclass
class DomStorageItemAdded:
    storage_id: StorageId
    key: str
    new_value: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemAdded:
        return cls(
            storage_id=StorageId.from_json(json['storageId']),
            key=str(json['key']),
            new_value=str(json['newValue'])
        )


@event_class('DOMStorage.domStorageItemRemoved')
@dataclass
class DomStorageItemRemoved:
    storage_id: StorageId
    key: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemRemoved:
        return cls(
            storage_id=StorageId.from_json(json['storageId']),
            key=str(json['key'])
        )


@event_class('DOMStorage.domStorageItemUpdated')
@dataclass
class DomStorageItemUpdated:
    storage_id: StorageId
    key: str
    old_value: str
    new_value: str

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemUpdated:
        return cls(
            storage_id=StorageId.from_json(json['storageId']),
            key=str(json['key']),
            old_value=str(json['oldValue']),
            new_value=str(json['newValue'])
        )


@event_class('DOMStorage.domStorageItemsCleared')
@dataclass
class DomStorageItemsCleared:
    storage_id: StorageId

    @classmethod
    def from_json(cls, json: T_JSON_DICT) -> DomStorageItemsCleared:
        return cls(
            storage_id=StorageId.from_json(json['storageId'])
        )